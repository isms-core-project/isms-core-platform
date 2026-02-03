#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.10.3 - Media Lifecycle Tracking Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.10: Storage Media
Assessment Domain 3 of 4: Media Lifecycle Management, Disposal & Re-use

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates an Excel assessment workbook for evaluating media lifecycle
management including acquisition, internal re-use, disposal methods, third-party
disposal, and paper document handling.

**Generated Workbook Structure:**
1. Instructions & Legend
2. Acquisition & Procurement
3. Internal Re-use Procedures
4. Disposal Methods
5. Third-Party Disposal
6. Paper Document Lifecycle
7. Summary Dashboard
8. Evidence Register
9. Approval Sign-Off

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

    python3 generate_a710_3_media_lifecycle.py

================================================================================
"""

# =============================================================================
# Imports
# =============================================================================
import logging
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.10.3"
WORKBOOK_NAME = "Media Lifecycle Tracking"
CONTROL_ID = "A.7.10"
CONTROL_NAME = "Storage Media"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
STYLES = {
    'title': {
        'font': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    },
    'header': {
        'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    },
    'subheader': {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D8E4F8', end_color='D8E4F8', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    },
    'input_cell': {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    }
}


def apply_style(cell, style_name):
    """Apply a named style to a cell."""
    style = STYLES[style_name]
    cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    cell.alignment = style['alignment']
    if 'border' in style:
        cell.border = style['border']


def create_workbook():
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions & Legend",
        "2. Acquisition & Procurement",
        "3. Internal Re-use",
        "4. Disposal Methods",
        "5. Third-Party Disposal",
        "6. Paper Document Lifecycle",
        "Summary Dashboard",
        "Evidence Register",
        "Approval Sign-Off"
    ]

    for idx, sheet_name in enumerate(sheets):
        wb.create_sheet(sheet_name, idx)

    return wb


def get_base_columns():
    """Return base column structure."""
    return [
        ("A", "Process / Category", 25),
        ("B", "Classification Level", 18),
        ("C", "Process Owner", 18),
        ("D", "Procedure Reference", 20),
        ("E", "Last Review Date", 12),
        ("F", "Status", 18),
        ("G", "Implementation Date", 12),
        ("H", "Last Audit Date", 12),
        ("I", "Next Review Date", 12),
        ("J", "Gap Identified", 25),
        ("K", "Remediation Plan", 25),
        ("L", "Target Completion", 12),
        ("M", "Risk Level", 12),
        ("N", "Evidence Reference", 20),
        ("O", "Notes / Comments", 25),
        ("P", "Remediation Owner", 18),
        ("Q", "Budget Required", 15)
    ]


def create_base_validations(ws):
    """Create data validation objects for standard columns."""
    dv_classification = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Restricted,All Classifications"',
        allow_blank=True
    )
    ws.add_data_validation(dv_classification)
    dv_classification.add('B10:B100')

    dv_status = DataValidation(
        type="list",
        formula1='"Compliant,Partial,Non-Compliant,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add('F10:F100')

    dv_risk = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_risk)
    dv_risk.add('M10:M100')

    dv_budget = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=True
    )
    ws.add_data_validation(dv_budget)
    dv_budget.add('Q10:Q100')


def create_assessment_sheet(ws, sheet_title, sheet_objective, extended_cols):
    """Create a standardized assessment sheet."""
    ws.merge_cells('A1:T1')
    cell = ws['A1']
    cell.value = sheet_title
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:T2')
    ws['A2'].value = "Policy Reference: ISMS-POL-A.7.10 - Storage Media"
    ws['A2'].font = Font(name='Calibri', size=9, italic=True)

    ws.merge_cells('A4:T6')
    cell = ws['A4']
    cell.value = f"ASSESSMENT OBJECTIVE:\n{sheet_objective}"
    cell.font = Font(name='Calibri', size=11, bold=True)
    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[4].height = 50

    ws.merge_cells('A7:T7')
    ws['A7'].value = "Complete the yellow-highlighted cells below. Use dropdowns where provided."
    ws['A7'].font = Font(name='Calibri', size=9, italic=True, color='0000FF')

    base_cols = get_base_columns()
    all_cols = base_cols + extended_cols

    for col_letter, header_text, width in all_cols:
        cell = ws[f'{col_letter}9']
        cell.value = header_text
        apply_style(cell, 'header')
        ws.column_dimensions[col_letter].width = width

    for row in range(10, 23):
        for col_idx in range(1, len(all_cols) + 1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f'{col_letter}{row}']
            apply_style(cell, 'input_cell')

    create_base_validations(ws)
    ws.freeze_panes = 'A10'


def create_instructions_sheet(ws):
    """Create Instructions & Legend sheet."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = f"{DOCUMENT_ID} - {WORKBOOK_NAME} Assessment"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    current_row = 3
    info_data = [
        ("Document ID:", DOCUMENT_ID),
        ("Assessment Date:", GENERATED_DATE),
        ("Assessor Name:", "[Enter Name]"),
        ("Organisation:", "[Enter Organisation]"),
        ("Control Reference:", CONTROL_REF)
    ]

    for label, value in info_data:
        ws[f'A{current_row}'].value = label
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'].value = value
        if "[" in value:
            ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        current_row += 1

    current_row += 1
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = "ASSESSMENT SCOPE"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    scope_items = [
        "- Acquisition & Procurement: Approved suppliers, registration controls",
        "- Internal Re-use: Secure erasure, verification, re-assignment",
        "- Disposal Methods: Destruction by classification level",
        "- Third-Party Disposal: Vendor management, certificates",
        "- Paper Document Lifecycle: Shredding standards, collection"
    ]

    for item in scope_items:
        ws[f'A{current_row}'].value = item
        current_row += 1

    current_row += 1
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = "DISPOSAL REQUIREMENTS BY CLASSIFICATION"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    disposal_reqs = [
        ("CONFIDENTIAL:", "Physical destruction with certificate"),
        ("INTERNAL:", "Secure overwriting (3+ passes) OR destruction"),
        ("PUBLIC:", "Standard deletion acceptable")
    ]

    for classification, requirement in disposal_reqs:
        ws[f'A{current_row}'].value = classification
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'].value = requirement
        current_row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60


def create_acquisition_sheet(ws):
    """Create Acquisition & Procurement sheet."""
    extended_cols = [
        ("R", "Approved Suppliers", 25),
        ("S", "Registration Required", 18),
        ("T", "Quality Standards", 22)
    ]

    create_assessment_sheet(
        ws,
        "2. Acquisition & Procurement",
        "Assess media procurement and registration controls including approved suppliers, encryption requirements, and asset registration processes.",
        extended_cols
    )

    dv_registration = DataValidation(
        type="list",
        formula1='"Yes - Automatic (PO),Yes - Manual Entry,No - Track Only,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_registration)
    dv_registration.add('S10:S100')


def create_reuse_sheet(ws):
    """Create Internal Re-use Procedures sheet."""
    extended_cols = [
        ("R", "Erasure Method", 25),
        ("S", "Verification Required", 20),
        ("T", "Documentation", 22)
    ]

    create_assessment_sheet(
        ws,
        "3. Internal Re-use Procedures",
        "Assess procedures for media re-assignment within the organisation including erasure methods, verification requirements, and documentation.",
        extended_cols
    )

    dv_erasure = DataValidation(
        type="list",
        formula1='"NIST 800-88 Clear,NIST 800-88 Purge,Cryptographic Erasure,Factory Reset,Full Format,Quick Format,Physical Destruction"',
        allow_blank=True
    )
    ws.add_data_validation(dv_erasure)
    dv_erasure.add('R10:R100')

    dv_verification = DataValidation(
        type="list",
        formula1='"Full Verification,Sample Verification (10%),Log Review Only,None"',
        allow_blank=True
    )
    ws.add_data_validation(dv_verification)
    dv_verification.add('S10:S100')

    dv_documentation = DataValidation(
        type="list",
        formula1='"Erasure Certificate,Asset System Log,Manual Log Entry,None Required"',
        allow_blank=True
    )
    ws.add_data_validation(dv_documentation)
    dv_documentation.add('T10:T100')


def create_disposal_methods_sheet(ws):
    """Create Disposal Methods sheet."""
    extended_cols = [
        ("R", "Destruction Method", 25),
        ("S", "Witness Required", 18),
        ("T", "Certificate Required", 18)
    ]

    create_assessment_sheet(
        ws,
        "4. Disposal Methods",
        "Assess disposal methods by classification level including destruction techniques, witness requirements, and certificate of destruction processes.",
        extended_cols
    )

    dv_destruction = DataValidation(
        type="list",
        formula1='"Physical Shredding,Degaussing,Degaussing + Shredding,Incineration,Secure Overwriting,Cryptographic Erasure,Crushing,Disintegration"',
        allow_blank=True
    )
    ws.add_data_validation(dv_destruction)
    dv_destruction.add('R10:R100')

    dv_witness = DataValidation(
        type="list",
        formula1='"Yes - Internal Witness,Yes - External Auditor,Yes - Both,No,For Certain Classifications"',
        allow_blank=True
    )
    ws.add_data_validation(dv_witness)
    dv_witness.add('S10:S100')

    dv_certificate = DataValidation(
        type="list",
        formula1='"Yes - Individual Certificate,Yes - Batch Certificate,No Certificate,Vendor Report Only"',
        allow_blank=True
    )
    ws.add_data_validation(dv_certificate)
    dv_certificate.add('T10:T100')


def create_third_party_sheet(ws):
    """Create Third-Party Disposal sheet."""
    extended_cols = [
        ("R", "Contract Status", 20),
        ("S", "Certifications", 25),
        ("T", "Certificate SLA", 18)
    ]

    create_assessment_sheet(
        ws,
        "5. Third-Party Disposal",
        "Assess third-party disposal vendor management including contract status, certifications (ISO 27001, NAID, R2, e-Stewards), and certificate delivery SLAs.",
        extended_cols
    )

    dv_contract = DataValidation(
        type="list",
        formula1='"Active Contract,Contract Pending Renewal,Contract Expired,No Contract (Spot Purchase),Under Evaluation"',
        allow_blank=True
    )
    ws.add_data_validation(dv_contract)
    dv_contract.add('R10:R100')


def create_paper_lifecycle_sheet(ws):
    """Create Paper Document Lifecycle sheet."""
    extended_cols = [
        ("R", "Shredding Standard", 22),
        ("S", "Collection Process", 22),
        ("T", "Destruction Frequency", 18)
    ]

    create_assessment_sheet(
        ws,
        "6. Paper Document Lifecycle",
        "Assess physical document handling and destruction including shredding standards (DIN 66399), collection processes, and destruction frequency.",
        extended_cols
    )

    dv_shredding = DataValidation(
        type="list",
        formula1='"DIN 66399 P-4 (Cross-cut),DIN 66399 P-5 (Fine Cross-cut),DIN 66399 P-6 (High Security),DIN 66399 P-7 (Super High Security),Strip Shred,Not Specified"',
        allow_blank=True
    )
    ws.add_data_validation(dv_shredding)
    dv_shredding.add('R10:R100')

    dv_collection = DataValidation(
        type="list",
        formula1='"Secure Bins (Locked),Secure Bins (Unlocked),On-Demand Shredding,Contractor Collection,Centralised Collection Point"',
        allow_blank=True
    )
    ws.add_data_validation(dv_collection)
    dv_collection.add('S10:S100')

    dv_frequency = DataValidation(
        type="list",
        formula1='"Daily,Weekly,Bi-weekly,Monthly,Quarterly,On-Demand,Event-Based"',
        allow_blank=True
    )
    ws.add_data_validation(dv_frequency)
    dv_frequency.add('T10:T100')


def create_summary_dashboard(ws):
    """Create Summary Dashboard sheet."""
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "Media Lifecycle Tracking - Summary Dashboard"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    current_row = 3
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "LIFECYCLE METRICS"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    metrics = [
        ("Media Acquired (This Period)", "[Count]"),
        ("Media Re-used (This Period)", "[Count]"),
        ("Media Disposed (This Period)", "[Count]"),
        ("Pending Disposal", "[Count]")
    ]

    for label, value in metrics:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        current_row += 1

    current_row += 1
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "DISPOSAL COMPLIANCE"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    disposal_metrics = [
        ("Certificates Received (%)", "[%]"),
        ("Witnessed Destructions (%)", "[%]"),
        ("Vendor Compliance (%)", "[%]"),
        ("Chain of Custody Complete (%)", "[%]")
    ]

    for label, value in disposal_metrics:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        current_row += 1

    current_row += 1
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "VENDOR STATUS"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    vendor_metrics = [
        ("Active Vendors", "[Count]"),
        ("Expiring Certifications", "[Count]"),
        ("Pending Contract Renewals", "[Count]"),
        ("Performance Issues", "[Count]")
    ]

    for label, value in vendor_metrics:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        current_row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15


def create_evidence_register(ws):
    """Create Evidence Register sheet."""
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "Evidence Register - Supporting Documentation"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Evidence ID", 12),
        ("B", "Assessment Sheet", 20),
        ("C", "Related Process", 30),
        ("D", "Evidence Type", 20),
        ("E", "Evidence Description", 35),
        ("F", "File Location/Link", 40),
        ("G", "Date Collected", 12),
        ("H", "Retention Period", 15),
        ("I", "Next Review", 12),
        ("J", "Owner", 20),
        ("K", "Notes", 30)
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f'{col_letter}4']
        cell.value = header_text
        apply_style(cell, 'header')
        ws.column_dimensions[col_letter].width = width

    for row in range(5, 105):
        for col_letter, _, _ in headers:
            cell = ws[f'{col_letter}{row}']
            apply_style(cell, 'input_cell')

    dv_type = DataValidation(
        type="list",
        formula1='"Certificate of Destruction,Erasure Report,Vendor Contract,Vendor Certification,Audit Report,Procedure Document,Training Record,Screenshot,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add('D5:D104')

    ws.freeze_panes = 'A5'


def create_approval_signoff(ws):
    """Create Approval Sign-Off sheet."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "Three-Level Approval Workflow"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    current_row = 3
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = "DOCUMENT CONTROL"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    fields = [
        ("Assessment Period:", "[e.g., Q1 2026]"),
        ("Workbook Version:", "1.0"),
        ("Lifecycle Processes Assessed:", "[Count]"),
        ("Overall Compliance %:", "[%]"),
        ("Assessment Completed By:", "[Name, Date]")
    ]

    for label, value in fields:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        current_row += 1

    current_row += 2

    approvals = [
        ("LEVEL 1 APPROVAL - Technical/Operational", "IT Operations Manager / Asset Management Lead"),
        ("LEVEL 2 APPROVAL - Management", "CISO / IT Director"),
        ("LEVEL 3 APPROVAL - Executive", "COO / CRO / Executive Management")
    ]

    approval_fields = ["Approver Name:", "Title:", "Date:", "Status:", "Comments:", "Signature:"]

    for level_title, role in approvals:
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'].value = level_title
        apply_style(ws[f'A{current_row}'], 'subheader')
        current_row += 1

        ws[f'A{current_row}'].value = f"Role: {role}"
        ws[f'A{current_row}'].font = Font(italic=True)
        current_row += 1

        for field in approval_fields:
            ws[f'A{current_row}'].value = field
            ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            if field == "Status:":
                dv = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected"', allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(f'B{current_row}')
            current_row += 1

        current_row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50


def main():
    """Main function to generate the workbook."""
    logger.info(f"Starting generation of {DOCUMENT_ID} - {WORKBOOK_NAME}")

    try:
        wb = create_workbook()

        logger.info("Creating Instructions & Legend sheet...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("Creating Acquisition & Procurement sheet...")
        create_acquisition_sheet(wb["2. Acquisition & Procurement"])

        logger.info("Creating Internal Re-use sheet...")
        create_reuse_sheet(wb["3. Internal Re-use"])

        logger.info("Creating Disposal Methods sheet...")
        create_disposal_methods_sheet(wb["4. Disposal Methods"])

        logger.info("Creating Third-Party Disposal sheet...")
        create_third_party_sheet(wb["5. Third-Party Disposal"])

        logger.info("Creating Paper Document Lifecycle sheet...")
        create_paper_lifecycle_sheet(wb["6. Paper Document Lifecycle"])

        logger.info("Creating Summary Dashboard...")
        create_summary_dashboard(wb["Summary Dashboard"])

        logger.info("Creating Evidence Register...")
        create_evidence_register(wb["Evidence Register"])

        logger.info("Creating Approval Sign-Off...")
        create_approval_signoff(wb["Approval Sign-Off"])

        logger.info(f"Saving workbook as {OUTPUT_FILENAME}...")
        wb.save(OUTPUT_FILENAME)

        logger.info("=" * 60)
        logger.info(f"SUCCESS: Workbook generated successfully!")
        logger.info(f"Output file: {OUTPUT_FILENAME}")
        logger.info("=" * 60)

        return True

    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation following A.8.10 pattern
# =============================================================================
