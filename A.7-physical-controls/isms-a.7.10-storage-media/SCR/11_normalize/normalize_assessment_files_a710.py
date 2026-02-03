#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
# =============================================================================
"""
================================================================================
ISMS-A.7.10 - Assessment File Normalizer Utility
================================================================================

ISO/IEC 27001:2022 Control A.7.10: Storage Media
Quality Assurance Utility: Excel Assessment File Normalization & Validation

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This utility normalizes assessment workbook filenames to standardized names
required for external workbook formula linking in the A.7.10.4 Compliance
Dashboard.

**Purpose:**
The A.7.10.4 Compliance Dashboard uses external workbook formulas like:
```
='[ISMS-IMP-A.7.10.1.xlsx]Summary Dashboard'!$G$10
```
These formulas require EXACT filenames. This script copies completed assessments
to normalized names, enabling external formula linking to work correctly.

**Normalization Scope:**
- ISMS-IMP-A.7.10.1_*.xlsx -> ISMS-IMP-A.7.10.1.xlsx
- ISMS-IMP-A.7.10.2_*.xlsx -> ISMS-IMP-A.7.10.2.xlsx
- ISMS-IMP-A.7.10.3_*.xlsx -> ISMS-IMP-A.7.10.3.xlsx

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

    python3 normalize_assessment_files_a710.py [--source DIR] [--output DIR] [-y]

================================================================================
"""

# =============================================================================
# Imports
# =============================================================================
import logging
import sys
import os
import shutil
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed")
    print("Install with: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# Configuration
# =============================================================================
EXPECTED_DOCS = {
    "ISMS-IMP-A.7.10.1": {
        "title": "Storage Media Inventory",
        "normalized": "ISMS-IMP-A.7.10.1.xlsx"
    },
    "ISMS-IMP-A.7.10.2": {
        "title": "Media Handling Procedures",
        "normalized": "ISMS-IMP-A.7.10.2.xlsx"
    },
    "ISMS-IMP-A.7.10.3": {
        "title": "Media Lifecycle Tracking",
        "normalized": "ISMS-IMP-A.7.10.3.xlsx"
    }
}


def validate_workbook(filepath):
    """
    Validate workbook contains expected document ID in Instructions sheet.

    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        if "Instructions & Legend" not in wb.sheetnames:
            wb.close()
            return (None, None)

        ws = wb["Instructions & Legend"]

        for row in range(3, 25):
            cell_label = ws.cell(row=row, column=1).value

            if cell_label and "Document ID" in str(cell_label):
                doc_id_value = ws.cell(row=row, column=2).value

                if doc_id_value:
                    doc_id = str(doc_id_value).strip()

                    if doc_id in EXPECTED_DOCS:
                        title = EXPECTED_DOCS[doc_id]["title"]
                        wb.close()
                        return (doc_id, title)

        wb.close()
        return (None, None)

    except Exception as e:
        logger.warning(f"Error reading file: {e}")
        return (None, None)


def get_file_info(filepath):
    """Get file metadata for manifest."""
    stat = filepath.stat()
    return {
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime),
    }


def scan_directory(directory):
    """Scan directory for assessment workbooks."""
    found_assessments = {}
    directory = Path(directory)

    xlsx_files = [f for f in directory.glob("*.xlsx") if not f.name.startswith("~$")]

    if not xlsx_files:
        logger.warning(f"No Excel files found in {directory}")
        return found_assessments

    logger.info(f"Scanning {len(xlsx_files)} Excel file(s) in {directory}...")

    for filepath in sorted(xlsx_files):
        logger.info(f"Checking: {filepath.name}")
        doc_id, title = validate_workbook(filepath)

        if doc_id:
            logger.info(f"  Valid: {doc_id} - {title}")
            found_assessments[doc_id] = {
                'path': filepath,
                'info': get_file_info(filepath)
            }
        else:
            logger.info(f"  Skipped (not a valid A.7.10 assessment workbook)")

    return found_assessments


def create_manifest(output_dir, mapping, source_dir):
    """Create audit manifest documenting the normalization process."""
    manifest_path = Path(output_dir) / "normalization_manifest_a710.txt"

    with open(manifest_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("ISMS ASSESSMENT FILE NORMALIZATION MANIFEST\n")
        f.write("ISO/IEC 27001:2022 - Control A.7.10: Storage Media\n")
        f.write("=" * 80 + "\n\n")

        f.write("NORMALIZATION METADATA\n")
        f.write("-" * 80 + "\n")
        f.write(f"Normalization Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Source Directory:        {source_dir.resolve()}\n")
        f.write(f"Output Directory:        {output_dir.resolve()}\n")
        f.write(f"Files Normalized:        {len(mapping)}/3 required\n")
        f.write(f"Normalization Status:    {'COMPLETE' if len(mapping) == 3 else 'INCOMPLETE'}\n")
        f.write("\n")

        f.write("FILE MAPPING\n")
        f.write("-" * 80 + "\n\n")

        for doc_id in sorted(EXPECTED_DOCS.keys()):
            if doc_id in mapping:
                info = mapping[doc_id]
                source_path = info['path']
                normalized_name = info['normalized']
                metadata = info['info']

                f.write(f"Document ID:      {doc_id}\n")
                f.write(f"Title:            {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"Source File:      {source_path.name}\n")
                f.write(f"Normalized Name:  {normalized_name}\n")
                f.write(f"File Size:        {metadata['size']:,} bytes\n")
                f.write(f"Status:           NORMALIZED\n")
                f.write("\n")
            else:
                f.write(f"Document ID:      {doc_id}\n")
                f.write(f"Title:            {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"Status:           NOT FOUND\n")
                f.write("\n")

        f.write("=" * 80 + "\n")
        f.write("END OF MANIFEST\n")
        f.write("=" * 80 + "\n")

    return manifest_path


def normalize_files(source_dir=None, output_dir=None, auto_confirm=False):
    """Main normalization orchestration function."""
    logger.info("=" * 60)
    logger.info("ISMS Assessment File Normalization Utility")
    logger.info("ISO/IEC 27001:2022 - Control A.7.10: Storage Media")
    logger.info("=" * 60)

    # Get source directory
    if not source_dir:
        source_dir = input("Enter source directory (or press Enter for current): ").strip()
        if not source_dir:
            source_dir = "."

    source_dir = Path(source_dir).resolve()

    if not source_dir.exists() or not source_dir.is_dir():
        logger.error(f"Invalid source directory: {source_dir}")
        return False

    logger.info(f"Source directory: {source_dir}")

    # Get output directory
    if not output_dir:
        default_output = source_dir / "Dashboard_Sources"
        output_input = input(f"Enter output directory (or press Enter for '{default_output.name}'): ").strip()
        if not output_input:
            output_dir = default_output
        else:
            output_dir = Path(output_input)

    output_dir = Path(output_dir).resolve()

    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Output directory: {output_dir}")
    except Exception as e:
        logger.error(f"Error creating output directory: {e}")
        return False

    # Scan for assessment files
    found = scan_directory(source_dir)

    if not found:
        logger.error("No valid assessment workbooks found")
        return False

    # Show normalization summary
    logger.info(f"Found {len(found)} of {len(EXPECTED_DOCS)} required assessment workbooks")

    mapping = {}
    for doc_id in sorted(EXPECTED_DOCS.keys()):
        if doc_id in found:
            normalized_name = EXPECTED_DOCS[doc_id]["normalized"]
            source_path = found[doc_id]['path']

            logger.info(f"  {doc_id}: {source_path.name} -> {normalized_name}")

            mapping[doc_id] = {
                'path': source_path,
                'normalized': normalized_name,
                'info': found[doc_id]['info']
            }
        else:
            logger.warning(f"  {doc_id}: NOT FOUND")

    # Confirm normalization
    if not auto_confirm:
        response = input("Proceed with normalization? (y/n): ").strip().lower()
        if response != 'y':
            logger.info("Normalization cancelled by user")
            return False

    # Perform normalization
    logger.info("Normalizing files...")

    for doc_id, info in mapping.items():
        source = info['path']
        dest = output_dir / info['normalized']

        logger.info(f"Copying: {source.name} -> {dest.name}")

        try:
            shutil.copy2(source, dest)
            logger.info(f"  Success")
        except Exception as e:
            logger.error(f"  Error: {e}")
            return False

    # Create audit manifest
    logger.info("Creating audit manifest...")
    try:
        manifest = create_manifest(output_dir, mapping, source_dir)
        logger.info(f"  Created: {manifest.name}")
    except Exception as e:
        logger.error(f"  Error creating manifest: {e}")
        return False

    # Success summary
    logger.info("=" * 60)
    logger.info("NORMALIZATION COMPLETE")
    logger.info("=" * 60)
    logger.info(f"Normalized files: {output_dir}")
    logger.info(f"Audit manifest:   {manifest}")
    logger.info("")
    logger.info("Next steps:")
    logger.info("  1. Generate dashboard: python3 generate_a710_4_compliance_dashboard.py")
    logger.info("  2. Place dashboard in same directory as normalized files")
    logger.info("  3. Open dashboard and click 'Update Links' when prompted")

    return True


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Normalize ISMS A.7.10 assessment workbooks for dashboard integration"
    )

    parser.add_argument('--source', '-s', help='Source directory')
    parser.add_argument('--output', '-o', help='Output directory')
    parser.add_argument('--auto-confirm', '-y', action='store_true', help='Skip confirmation')

    args = parser.parse_args()

    success = normalize_files(
        source_dir=args.source,
        output_dir=args.output,
        auto_confirm=args.auto_confirm
    )

    sys.exit(0 if success else 1)


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation following A.8.10 pattern
# =============================================================================
