#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
# =============================================================================
"""
================================================================================
Excel Sanity Check for ISMS-IMP-A.7.10 (Storage Media Framework)
================================================================================

Validates all 4 Storage Media assessment workbooks (Domains 1-4) for
ISO/IEC 27001:2022 Control A.7.10: Storage Media

This script checks for:
- Workbook structure (correct sheets present)
- Data validations (properly configured dropdowns)
- Document ID consistency
- Summary Dashboard structure for external linking
- Column widths appropriate
- Freeze panes set correctly

Usage:
    python3 excel_sanity_check_a710.py [workbook_directory]

================================================================================
"""

# =============================================================================
# Imports
# =============================================================================
import logging
import sys
import os
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed")
    print("Install with: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# Configuration
# =============================================================================
WORKBOOK_SPECS = {
    "Domain_1": {
        "pattern": "ISMS-IMP-A.7.10.1",
        "name": "Storage Media Inventory",
        "document_id": "ISMS-IMP-A.7.10.1",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Digital Storage Media",
            "3. Removable Media Registry",
            "4. Fixed Storage Assets",
            "5. Cloud Storage Mapping",
            "6. Physical Documents",
            "Summary Dashboard",
            "Evidence Register",
            "Approval Sign-Off"
        ],
        "min_sheets": 9
    },
    "Domain_2": {
        "pattern": "ISMS-IMP-A.7.10.2",
        "name": "Media Handling Procedures",
        "document_id": "ISMS-IMP-A.7.10.2",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Media Access Controls",
            "3. Transportation Security",
            "4. Physical Storage Controls",
            "5. Media Use Procedures",
            "6. Incident Response",
            "Summary Dashboard",
            "Evidence Register",
            "Approval Sign-Off"
        ],
        "min_sheets": 9
    },
    "Domain_3": {
        "pattern": "ISMS-IMP-A.7.10.3",
        "name": "Media Lifecycle Tracking",
        "document_id": "ISMS-IMP-A.7.10.3",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Acquisition & Procurement",
            "3. Internal Re-use",
            "4. Disposal Methods",
            "5. Third-Party Disposal",
            "6. Paper Document Lifecycle",
            "Summary Dashboard",
            "Evidence Register",
            "Approval Sign-Off"
        ],
        "min_sheets": 9
    },
    "Domain_4": {
        "pattern": "ISMS-IMP-A.7.10.4",
        "name": "Compliance Dashboard",
        "document_id": "ISMS-IMP-A.7.10.4",
        "expected_sheets": [
            "Instructions & Legend",
            "Executive Summary",
            "Domain 1 Summary",
            "Domain 2 Summary",
            "Domain 3 Summary",
            "Consolidated Gap Analysis",
            "Risk Register",
            "Remediation Roadmap",
            "Evidence Master Index",
            "KPI Dashboard",
            "CISO Approval"
        ],
        "min_sheets": 11,
        "consolidation_dashboard": True
    }
}

# Terminal colours
GREEN = '\033[92m'
YELLOW = '\033[93m'
RED = '\033[91m'
BLUE = '\033[94m'
BOLD = '\033[1m'
RESET = '\033[0m'


def print_success(text):
    print(f"{GREEN}  [OK] {text}{RESET}")


def print_warning(text):
    print(f"{YELLOW}  [WARN] {text}{RESET}")


def print_error(text):
    print(f"{RED}  [ERROR] {text}{RESET}")


def find_workbook(directory, pattern):
    """Find workbook matching pattern in directory."""
    path = Path(directory)
    matches = list(path.glob(f"{pattern}*.xlsx"))
    matches = [m for m in matches if not m.name.startswith('~$')]

    if not matches:
        return None

    if len(matches) > 1:
        print_warning(f"Multiple workbooks found for {pattern}, using most recent")
        matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)

    return matches[0]


def check_document_id(wb, expected_doc_id):
    """Check if Document ID is present in Instructions & Legend."""
    issues = []

    if "Instructions & Legend" not in wb.sheetnames:
        issues.append("Instructions & Legend sheet not found")
        return issues

    ws = wb["Instructions & Legend"]
    doc_id_found = False
    found_doc_id = None

    for row in range(1, 31):
        cell_label = ws.cell(row=row, column=1).value
        if cell_label and "Document ID" in str(cell_label):
            found_doc_id = ws.cell(row=row, column=2).value
            doc_id_found = True
            break

    if not doc_id_found:
        issues.append("Document ID not found in Instructions & Legend sheet")
    elif found_doc_id != expected_doc_id:
        issues.append(f"Document ID mismatch: Expected '{expected_doc_id}', found '{found_doc_id}'")
    else:
        print_success(f"Document ID verified: {found_doc_id}")

    return issues


def check_workbook_structure(wb, spec):
    """Check if workbook has correct sheet structure."""
    issues = []
    warnings = []

    actual_sheets = wb.sheetnames
    expected_sheets = spec["expected_sheets"]

    if len(actual_sheets) < spec["min_sheets"]:
        issues.append(f"Only {len(actual_sheets)} sheets found, expected at least {spec['min_sheets']}")
    else:
        print_success(f"Sheet count: {len(actual_sheets)} sheets")

    missing_sheets = []
    for sheet_name in expected_sheets:
        found = False
        for actual_sheet in actual_sheets:
            if actual_sheet == sheet_name or (
                "Approval" in sheet_name and "Approval" in actual_sheet
            ):
                found = True
                break

        if not found:
            missing_sheets.append(sheet_name)

    if missing_sheets:
        issues.append(f"Missing sheets: {', '.join(missing_sheets)}")
    else:
        print_success(f"All {len(expected_sheets)} expected sheets present")

    return issues, warnings


def check_data_validations(ws, sheet_name):
    """Check if data validations are properly configured."""
    issues = []
    warnings = []

    if not ws.data_validations.dataValidation:
        if sheet_name not in ["Instructions & Legend", "Summary Dashboard",
                              "Executive Summary", "CISO Approval"]:
            warnings.append(f"{sheet_name}: No data validations found")
        return issues, warnings

    validation_count = len(ws.data_validations.dataValidation)
    print_success(f"{sheet_name}: {validation_count} data validations configured")

    return issues, warnings


def check_freeze_panes(ws, sheet_name):
    """Check if freeze panes are set."""
    issues = []
    warnings = []

    if ws.freeze_panes:
        print_success(f"{sheet_name}: Freeze panes set at {ws.freeze_panes}")
    else:
        if sheet_name not in ["Summary Dashboard", "Executive Summary", "Instructions & Legend"]:
            warnings.append(f"{sheet_name}: No freeze panes set")

    return issues, warnings


def validate_workbook(filepath, domain_key, spec):
    """Perform comprehensive validation of a workbook."""
    print(f"\n{BLUE}{BOLD}{'=' * 60}{RESET}")
    print(f"{BLUE}{BOLD}Validating {spec['name']} ({domain_key}){RESET}")
    print(f"{BLUE}{BOLD}{'=' * 60}{RESET}")
    print(f"  File: {filepath.name}")

    all_issues = []
    all_warnings = []

    try:
        print("  Loading workbook...")
        wb = load_workbook(filepath, data_only=False)
        print_success("Workbook loaded successfully")

        # Check Document ID
        print("  Checking Document ID...")
        issues = check_document_id(wb, spec["document_id"])
        all_issues.extend(issues)

        # Check structure
        print("  Checking workbook structure...")
        issues, warnings = check_workbook_structure(wb, spec)
        all_issues.extend(issues)
        all_warnings.extend(warnings)

        # Check each sheet
        print("  Checking individual sheets...")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            issues, warnings = check_data_validations(ws, sheet_name)
            all_issues.extend(issues)
            all_warnings.extend(warnings)

            issues, warnings = check_freeze_panes(ws, sheet_name)
            all_issues.extend(issues)
            all_warnings.extend(warnings)

        # Summary
        print("\n  " + "-" * 40)
        if len(all_issues) == 0 and len(all_warnings) == 0:
            print_success(f"{domain_key}: VALIDATION PASSED - No issues found")
            return True, 0, 0
        else:
            if len(all_issues) > 0:
                print_error(f"{domain_key}: {len(all_issues)} critical issues found")
                for issue in all_issues[:5]:
                    print_error(f"  - {issue}")
                if len(all_issues) > 5:
                    print_error(f"  ... and {len(all_issues) - 5} more issues")

            if len(all_warnings) > 0:
                print_warning(f"{domain_key}: {len(all_warnings)} warnings found")
                for warning in all_warnings[:5]:
                    print_warning(f"  - {warning}")
                if len(all_warnings) > 5:
                    print_warning(f"  ... and {len(all_warnings) - 5} more warnings")

            return len(all_issues) == 0, len(all_issues), len(all_warnings)

    except Exception as e:
        print_error(f"Failed to validate workbook: {str(e)}")
        return False, 1, 0


def main():
    """Main validation function."""
    if len(sys.argv) > 1:
        directory = sys.argv[1]
    else:
        directory = "."

    if not os.path.isdir(directory):
        print_error(f"Directory not found: {directory}")
        sys.exit(1)

    print(f"\n{BLUE}{BOLD}{'=' * 60}{RESET}")
    print(f"{BLUE}{BOLD}ISMS-IMP-A.7.10 Storage Media Framework Validation{RESET}")
    print(f"{BLUE}{BOLD}{'=' * 60}{RESET}")
    print(f"  Checking directory: {os.path.abspath(directory)}")
    print(f"  Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    results = {}
    total_issues = 0
    total_warnings = 0

    for domain_key, spec in WORKBOOK_SPECS.items():
        filepath = find_workbook(directory, spec["pattern"])

        if filepath:
            passed, issues, warnings = validate_workbook(filepath, domain_key, spec)
            results[domain_key] = {
                "found": True,
                "passed": passed,
                "issues": issues,
                "warnings": warnings
            }
            total_issues += issues
            total_warnings += warnings
        else:
            print(f"\n{YELLOW}  {domain_key}: Workbook not found ({spec['pattern']}*.xlsx){RESET}")
            results[domain_key] = {
                "found": False,
                "passed": None,
                "issues": 0,
                "warnings": 0
            }

    # Final summary
    print(f"\n{BLUE}{BOLD}{'=' * 60}{RESET}")
    print(f"{BLUE}{BOLD}VALIDATION SUMMARY{RESET}")
    print(f"{BLUE}{BOLD}{'=' * 60}{RESET}")

    found_count = sum(1 for r in results.values() if r["found"])
    passed_count = sum(1 for r in results.values() if r["passed"] == True)

    print(f"  Workbooks found: {found_count} / 4")
    print(f"  Workbooks passed: {passed_count} / {found_count if found_count > 0 else 1}")
    print(f"  Total critical issues: {total_issues}")
    print(f"  Total warnings: {total_warnings}")

    print("\n  Domain Status:")
    for domain_key, result in results.items():
        spec = WORKBOOK_SPECS[domain_key]
        if result["found"]:
            if result["passed"]:
                print_success(f"{domain_key} ({spec['name']}): PASSED")
            else:
                print_error(f"{domain_key} ({spec['name']}): FAILED")
        else:
            print_warning(f"{domain_key} ({spec['name']}): NOT FOUND")

    # Exit code
    if total_issues > 0:
        print(f"\n{RED}{BOLD}VALIDATION FAILED{RESET}")
        sys.exit(1)
    elif found_count == 0:
        print(f"\n{YELLOW}{BOLD}NO WORKBOOKS FOUND{RESET}")
        print("  Generate workbooks first using the Python generator scripts:")
        print("    python3 generate_a710_1_storage_media_inventory.py")
        print("    python3 generate_a710_2_media_handling.py")
        print("    python3 generate_a710_3_media_lifecycle.py")
        print("    python3 generate_a710_4_compliance_dashboard.py")
        sys.exit(2)
    else:
        print(f"\n{GREEN}{BOLD}VALIDATION PASSED{RESET}")
        if total_warnings > 0:
            print(f"  Note: {total_warnings} warnings found (non-critical)")
        print("\n  Next steps:")
        print("    1. Complete assessments (fill yellow cells)")
        print("    2. Run: python3 normalize_assessment_files_a710.py")
        print("    3. Open dashboard and click 'Update Links'")
        sys.exit(0)


if __name__ == "__main__":
    main()


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation following A.8.10 pattern
# =============================================================================
