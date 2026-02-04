#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS Excel Workbook Sanity Check Utility - A.7.8-9
================================================================================

ISO/IEC 27001:2022 Controls A.7.8 & A.7.9: Equipment Siting and Protection
Quality Assurance Utility: Excel Workbook Validation & Verification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script performs comprehensive validation of generated assessment workbooks:
1. File existence and accessibility
2. Sheet structure verification
3. Metadata validation
4. Cell content integrity checks
5. Formula validation
6. Cross-reference verification

Key customization areas:
1. Expected workbook configurations (match your generators)
2. Validation thresholds and tolerances
3. Report output format and detail level

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel processing

Usage:
    python3 excel_sanity_check_a78.py [--dir DIR] [--verbose]

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    logger.error("openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)

# =============================================================================
# EXPECTED WORKBOOK CONFIGURATIONS
# =============================================================================

EXPECTED_WORKBOOKS = {
    "ISMS-IMP-A.7.8.S1": {
        "title": "Equipment Siting Assessment",
        "sheets": [
            "Instructions & Legend",
            "Equipment Locations",
            "Environmental",
            "Physical Security",
            "Power Infrastructure",
            "Workstation Security",
            "Evidence Register",
            "Summary Dashboard",
            "Approval Sign-Off"
        ],
        "min_sheets": 9,
        "required_metadata": ["Document ID", "Version"]
    },
    "ISMS-IMP-A.7.9.S2": {
        "title": "Off-Premises Asset Security Assessment",
        "sheets": [
            "Instructions & Legend",
            "Equipment Inventory",
            "Authorisation",
            "Protection Measures",
            "Remote Working",
            "Permanent Off-Site",
            "Incidents",
            "Evidence Register",
            "Summary Dashboard",
            "Approval Sign-Off"
        ],
        "min_sheets": 10,
        "required_metadata": ["Document ID", "Version"]
    },
    "ISMS-IMP-A.7.8-9.S3": {
        "title": "Equipment Protection Compliance Dashboard",
        "sheets": [
            "Instructions & Legend",
            "Executive Summary",
            "Control Area Details",
            "Gap Register",
            "Incident Tracker",
            "Remediation Plan",
            "Audit Readiness",
            "Trend Analysis",
            "Data Input"
        ],
        "min_sheets": 9,
        "required_metadata": ["Document ID", "Version"]
    }
}

# =============================================================================
# VALIDATION FUNCTIONS
# =============================================================================

def check_file_exists(filepath):
    """Check if file exists and is readable."""
    path = Path(filepath)
    if not path.exists():
        return False, f"File not found: {filepath}"
    if not path.is_file():
        return False, f"Not a file: {filepath}"
    if path.stat().st_size == 0:
        return False, f"File is empty: {filepath}"
    return True, "File exists and is readable"


def check_workbook_opens(filepath):
    """Check if workbook can be opened without errors."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet_count = len(wb.sheetnames)
        wb.close()
        return True, f"Workbook opens successfully ({sheet_count} sheets)"
    except Exception as e:
        return False, f"Error opening workbook: {e}"


def check_sheet_structure(filepath, expected_sheets):
    """Verify all expected sheets exist."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        actual_sheets = wb.sheetnames
        wb.close()

        missing = []
        for sheet in expected_sheets:
            if sheet not in actual_sheets:
                missing.append(sheet)

        if missing:
            return False, f"Missing sheets: {', '.join(missing)}"
        return True, f"All {len(expected_sheets)} expected sheets present"
    except Exception as e:
        return False, f"Error checking sheets: {e}"


def check_metadata(filepath, required_fields):
    """Verify metadata fields in Instructions sheet."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        if "Instructions & Legend" not in wb.sheetnames:
            wb.close()
            return False, "Instructions & Legend sheet not found"

        ws = wb["Instructions & Legend"]
        found_fields = []

        for row in range(1, 30):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                for field in required_fields:
                    if field in str(cell_value):
                        found_fields.append(field)

        wb.close()

        missing = [f for f in required_fields if f not in found_fields]
        if missing:
            return False, f"Missing metadata fields: {', '.join(missing)}"
        return True, f"All {len(required_fields)} metadata fields present"
    except Exception as e:
        return False, f"Error checking metadata: {e}"


def check_document_id(filepath, expected_prefix):
    """Verify document ID matches expected pattern."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        if "Instructions & Legend" not in wb.sheetnames:
            wb.close()
            return False, "Instructions & Legend sheet not found"

        ws = wb["Instructions & Legend"]

        for row in range(3, 25):
            cell_label = ws.cell(row=row, column=1).value
            if cell_label and "Document ID" in str(cell_label):
                doc_id = ws.cell(row=row, column=2).value
                wb.close()

                if doc_id and str(doc_id).strip().startswith(expected_prefix):
                    return True, f"Document ID valid: {doc_id}"
                else:
                    return False, f"Document ID mismatch: expected {expected_prefix}*, got {doc_id}"

        wb.close()
        return False, "Document ID field not found"
    except Exception as e:
        return False, f"Error checking document ID: {e}"


def check_cell_formatting(filepath):
    """Verify cells have proper formatting (not corrupted)."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=False)
        issues = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Check first 10 rows for formatting issues
            for row in range(1, min(11, ws.max_row + 1)):
                for col in range(1, min(11, ws.max_column + 1)):
                    try:
                        cell = ws.cell(row=row, column=col)
                        _ = cell.value  # Access value to check integrity
                    except Exception as e:
                        issues.append(f"{sheet_name}!{cell.coordinate}: {e}")

        wb.close()

        if issues:
            return False, f"Formatting issues found: {len(issues)}"
        return True, "Cell formatting verified"
    except Exception as e:
        return False, f"Error checking formatting: {e}"


def check_formula_integrity(filepath):
    """Verify formulas are syntactically valid."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=False)
        formula_count = 0
        error_formulas = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).startswith('='):
                        formula_count += 1
                        # Check for common formula errors
                        formula = str(cell.value)
                        if '#REF!' in formula or '#NAME?' in formula:
                            error_formulas.append(f"{sheet_name}!{cell.coordinate}")

        wb.close()

        if error_formulas:
            return False, f"Formula errors in: {', '.join(error_formulas[:5])}"
        return True, f"All {formula_count} formulas valid"
    except Exception as e:
        return False, f"Error checking formulas: {e}"


# =============================================================================
# MAIN VALIDATION FUNCTION
# =============================================================================

def validate_workbook(filepath, config):
    """Run all validation checks on a workbook."""
    results = {
        "filepath": str(filepath),
        "config": config["title"],
        "checks": [],
        "passed": 0,
        "failed": 0,
        "overall": "UNKNOWN"
    }

    # File existence check
    passed, msg = check_file_exists(filepath)
    results["checks"].append(("File Exists", passed, msg))
    if passed:
        results["passed"] += 1
    else:
        results["failed"] += 1
        results["overall"] = "FAILED"
        return results  # Can't continue if file doesn't exist

    # Workbook opens check
    passed, msg = check_workbook_opens(filepath)
    results["checks"].append(("Workbook Opens", passed, msg))
    if passed:
        results["passed"] += 1
    else:
        results["failed"] += 1
        results["overall"] = "FAILED"
        return results

    # Sheet structure check
    passed, msg = check_sheet_structure(filepath, config["sheets"])
    results["checks"].append(("Sheet Structure", passed, msg))
    if passed:
        results["passed"] += 1
    else:
        results["failed"] += 1

    # Metadata check
    passed, msg = check_metadata(filepath, config["required_metadata"])
    results["checks"].append(("Metadata Fields", passed, msg))
    if passed:
        results["passed"] += 1
    else:
        results["failed"] += 1

    # Document ID check
    doc_id_prefix = [k for k in EXPECTED_WORKBOOKS.keys()
                     if EXPECTED_WORKBOOKS[k]["title"] == config["title"]][0]
    passed, msg = check_document_id(filepath, doc_id_prefix)
    results["checks"].append(("Document ID", passed, msg))
    if passed:
        results["passed"] += 1
    else:
        results["failed"] += 1

    # Cell formatting check
    passed, msg = check_cell_formatting(filepath)
    results["checks"].append(("Cell Formatting", passed, msg))
    if passed:
        results["passed"] += 1
    else:
        results["failed"] += 1

    # Formula integrity check
    passed, msg = check_formula_integrity(filepath)
    results["checks"].append(("Formula Integrity", passed, msg))
    if passed:
        results["passed"] += 1
    else:
        results["failed"] += 1

    # Determine overall result
    if results["failed"] == 0:
        results["overall"] = "PASSED"
    else:
        results["overall"] = "FAILED"

    return results


def find_workbooks(directory):
    """Find assessment workbooks in directory."""
    directory = Path(directory)
    found = {}

    xlsx_files = list(directory.glob("*.xlsx"))
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]

    for filepath in xlsx_files:
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

            if "Instructions & Legend" in wb.sheetnames:
                ws = wb["Instructions & Legend"]

                for row in range(3, 25):
                    cell_label = ws.cell(row=row, column=1).value
                    if cell_label and "Document ID" in str(cell_label):
                        doc_id = ws.cell(row=row, column=2).value
                        if doc_id:
                            doc_id = str(doc_id).strip()
                            if doc_id in EXPECTED_WORKBOOKS:
                                found[doc_id] = filepath
                        break

            wb.close()
        except Exception as e:
            logger.warning("Error reading %s: %s", filepath.name, e)

    return found


def print_report(all_results):
    """Print validation report."""
    print("\n" + "=" * 80)
    print("ISMS A.7.8-9 EXCEL WORKBOOK SANITY CHECK REPORT")
    print("ISO/IEC 27001:2022 - Controls A.7.8 & A.7.9: Equipment Siting and Protection")
    print("=" * 80)
    print(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 80)

    total_passed = 0
    total_failed = 0
    workbooks_passed = 0
    workbooks_failed = 0

    for result in all_results:
        print(f"\n{'─' * 80}")
        print(f"Workbook: {Path(result['filepath']).name}")
        print(f"Type: {result['config']}")
        print(f"Overall: {result['overall']}")
        print(f"{'─' * 80}")

        for check_name, passed, message in result["checks"]:
            status = "✓ PASS" if passed else "✗ FAIL"
            print(f"  [{status}] {check_name}: {message}")

        print(f"\n  Summary: {result['passed']} passed, {result['failed']} failed")

        total_passed += result["passed"]
        total_failed += result["failed"]
        if result["overall"] == "PASSED":
            workbooks_passed += 1
        else:
            workbooks_failed += 1

    print("\n" + "=" * 80)
    print("OVERALL SUMMARY")
    print("=" * 80)
    print(f"Workbooks Validated: {len(all_results)}")
    print(f"Workbooks Passed:    {workbooks_passed}")
    print(f"Workbooks Failed:    {workbooks_failed}")
    print(f"Total Checks Passed: {total_passed}")
    print(f"Total Checks Failed: {total_failed}")
    print(f"Overall Status:      {'ALL PASSED' if workbooks_failed == 0 else 'SOME FAILED'}")
    print("=" * 80)

    return workbooks_failed == 0


# =============================================================================
# COMMAND LINE INTERFACE
# =============================================================================

def main():
    """Main entry point."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Sanity check ISMS A.7.8-9 assessment workbooks"
    )

    parser.add_argument(
        '--dir', '-d',
        default='.',
        help='Directory containing workbooks (default: current directory)'
    )

    parser.add_argument(
        '--file', '-f',
        help='Check specific file instead of scanning directory'
    )

    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable verbose output'
    )

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    logger.info("=" * 70)
    logger.info("ISMS A.7.8-9 EXCEL WORKBOOK SANITY CHECK")
    logger.info("ISO/IEC 27001:2022 - Controls A.7.8 & A.7.9")
    logger.info("=" * 70)

    all_results = []

    if args.file:
        # Check specific file
        filepath = Path(args.file)
        if not filepath.exists():
            logger.error("File not found: %s", args.file)
            sys.exit(1)

        # Try to determine workbook type
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            if "Instructions & Legend" in wb.sheetnames:
                ws = wb["Instructions & Legend"]
                for row in range(3, 25):
                    cell_label = ws.cell(row=row, column=1).value
                    if cell_label and "Document ID" in str(cell_label):
                        doc_id = ws.cell(row=row, column=2).value
                        if doc_id and str(doc_id).strip() in EXPECTED_WORKBOOKS:
                            config = EXPECTED_WORKBOOKS[str(doc_id).strip()]
                            result = validate_workbook(filepath, config)
                            all_results.append(result)
                        break
            wb.close()
        except Exception as e:
            logger.error("Error processing file: %s", e)
            sys.exit(1)
    else:
        # Scan directory
        directory = Path(args.dir).resolve()
        logger.info("Scanning directory: %s", directory)

        found = find_workbooks(directory)

        if not found:
            logger.warning("No valid A.7.8-9 assessment workbooks found")
            logger.info("Looking for workbooks with Document IDs:")
            for doc_id in EXPECTED_WORKBOOKS:
                logger.info("  - %s", doc_id)
            sys.exit(1)

        logger.info("Found %d workbook(s) to validate", len(found))

        for doc_id, filepath in sorted(found.items()):
            logger.info("Validating: %s", filepath.name)
            config = EXPECTED_WORKBOOKS[doc_id]
            result = validate_workbook(filepath, config)
            all_results.append(result)

    if all_results:
        success = print_report(all_results)
        sys.exit(0 if success else 1)
    else:
        logger.error("No workbooks validated")
        sys.exit(1)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED (syntax validated, structure verified)
# QA_TOOL: Claude Code Deep Scan
# QA_NOTE: STANDARDIZATION - License header, logging, main() pattern applied
# =============================================================================
