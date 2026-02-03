#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.6.S1 - Secure Areas Working Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.6: Working in Secure Areas
Assessment Domain 1 of 4: Secure Area Procedures and Personnel Conduct

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates an assessment workbook for documenting secure areas,
working procedures, third-party access management, and incident tracking.

Reference: ISMS-IMP-A.7.6-7-14-S1 Specification

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    pip3 install openpyxl

Usage:
    python3 generate_a76_1_secure_areas.py

Output:
    ISMS-IMP-A.7.6.S1_Secure_Areas_Working_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.6.S1"
WORKBOOK_NAME = "Secure Areas Working Assessment"
CONTROL_ID = "A.7.6"
CONTROL_NAME = "Working in Secure Areas"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # Swiss format
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = '\u2705'      # Green checkmark
XMARK = '\u274C'      # Red X
WARNING = '\u26A0'    # Warning sign
BULLET = '\u2022'     # Bullet point

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Secure Areas Working Procedures"),
        ("Related Policy", "ISMS-POL-A.7.6-7-14"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "1. Document all defined secure areas in Sheet 2 (Secure Area Register).",
        "2. Assess working procedure compliance in Sheet 3 (Working Procedures).",
        "3. Track third-party access in Sheet 4 (Third-Party Access).",
        "4. Log security incidents in Sheet 5 (Incidents).",
        "5. Review Summary Dashboard (Sheet 6) for compliance metrics.",
        "6. Document evidence in Sheet 7 (Evidence Register).",
        "7. Obtain approvals in Sheet 8 (Approval Sign-Off).",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws[f"A{row+1}"] = "STATUS LEGEND"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    legend = [
        ("Symbol", "Status", "Description"),
        (f"{CHECK}", "Compliant", "Meets all policy requirements"),
        (f"{WARNING}", "Partial", "Meets some requirements, gaps identified"),
        (f"{XMARK}", "Non-Compliant", "Does not meet requirements"),
        ("N/A", "Not Applicable", "Not required for this area"),
    ]

    row += 2
    for c, h in enumerate(legend[0], start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for sym, status, desc in legend[1:]:
        ws.cell(row=row, column=1, value=sym).border = styles["border"]
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if "Compliant" in status and "Non" not in status:
            s.fill = styles["status_compliant"]["fill"]
        elif "Partial" in status:
            s.fill = styles["status_partial"]["fill"]
        elif "Non-Compliant" in status:
            s.fill = styles["status_noncompliant"]["fill"]
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50


def create_secure_area_register(ws, styles):
    """Create Secure Area Register sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "Secure Area Register\nPolicy Requirement: All secure areas must be defined, classified, and have appropriate controls"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Secure Area ID": 15,
        "Secure Area Name": 30,
        "Location": 25,
        "Classification": 20,
        "Access Control Type": 20,
        "Authorised Personnel": 18,
        "Last Access Review": 18,
        "Emergency Procedures": 18,
        "Recording Controls": 18,
        "Unsupervised Working": 18,
        "Status": 15,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data validation
    dv_classification = DataValidation(type="list", formula1='"Tier 1 - Critical,Tier 2 - Standard"', allow_blank=True)
    dv_access_type = DataValidation(type="list", formula1='"Badge + PIN,Biometric,Badge Only,Mantrap,Key Lock"', allow_blank=True)
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_recording = DataValidation(type="list", formula1='"Prohibited,Authorised Only,No Restriction"', allow_blank=True)
    dv_unsupervised = DataValidation(type="list", formula1='"Prohibited,Restricted,Allowed"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"', allow_blank=True)

    ws.add_data_validation(dv_classification)
    ws.add_data_validation(dv_access_type)
    ws.add_data_validation(dv_yesno)
    ws.add_data_validation(dv_recording)
    ws.add_data_validation(dv_unsupervised)
    ws.add_data_validation(dv_status)

    # Data rows
    for r in range(4, 104):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Apply data validations
        dv_classification.add(ws.cell(row=r, column=4))
        dv_access_type.add(ws.cell(row=r, column=5))
        dv_yesno.add(ws.cell(row=r, column=8))
        dv_recording.add(ws.cell(row=r, column=9))
        dv_unsupervised.add(ws.cell(row=r, column=10))
        dv_status.add(ws.cell(row=r, column=11))

    ws.freeze_panes = "A4"


def create_working_procedures(ws, styles):
    """Create Working Procedures assessment sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "Working Procedures Compliance Assessment\nPolicy Requirements per ISMS-POL-A.7.6-7-14 Section 2.1"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Requirement ID": 15,
        "Requirement Description": 50,
        "Implementation Status": 20,
        "Evidence": 40,
        "Status": 15,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate requirements
    requirements = [
        ("WP-001", "Secure area activities communicated on need-to-know basis"),
        ("WP-002", "Unsupervised working in secure areas avoided/restricted"),
        ("WP-003", "Empty secure areas locked and periodically checked"),
        ("WP-004", "Recording equipment (audio/video/photo) controlled"),
        ("WP-005", "Personal devices subject to strict controls in secure areas"),
        ("WP-006", "Emergency procedures accessible in secure areas"),
        ("WP-007", "Visitors escorted in secure areas at all times"),
        ("WP-008", "Visitor access logged and time-limited"),
        ("WP-009", "Third parties sign confidentiality agreements"),
        ("WP-010", "Third-party equipment inspected before entry"),
        ("WP-011", "Sensitive information not discussed where overheard"),
        ("WP-012", "Whiteboards/flipcharts erased after meetings"),
        ("WP-013", "Sensitive documents removed from printers immediately"),
        ("WP-014", "Clean desk requirements enforced in secure areas"),
    ]

    dv_impl = DataValidation(type="list", formula1='"Implemented,Partial,Not Implemented,N/A"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"', allow_blank=True)
    ws.add_data_validation(dv_impl)
    ws.add_data_validation(dv_status)

    row = 4
    for req_id, req_desc in requirements:
        ws.cell(row=row, column=1, value=req_id).border = styles["border"]
        ws.cell(row=row, column=2, value=req_desc).border = styles["border"]
        for c in range(3, 7):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_impl.add(ws.cell(row=row, column=3))
        dv_status.add(ws.cell(row=row, column=5))
        row += 1

    # Additional blank rows
    for r in range(row, row + 36):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_impl.add(ws.cell(row=r, column=3))
        dv_status.add(ws.cell(row=r, column=5))

    ws.freeze_panes = "A4"


def create_third_party_access(ws, styles):
    """Create Third-Party Access tracking sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Third-Party Access Tracking\nPolicy Requirement: All visitor/contractor access must be logged, escorted, and time-limited"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Access Period": 15,
        "Secure Area": 30,
        "Visitor/Contractor Count": 18,
        "Escort Compliance %": 18,
        "NDA Status": 15,
        "Time Limit Compliance": 18,
        "Equipment Inspected": 18,
        "Status": 15,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_yesno = DataValidation(type="list", formula1='"Yes,Partial,No,N/A"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"', allow_blank=True)
    ws.add_data_validation(dv_yesno)
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_yesno.add(ws.cell(row=r, column=5))
        dv_yesno.add(ws.cell(row=r, column=6))
        dv_yesno.add(ws.cell(row=r, column=7))
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"


def create_incidents_sheet(ws, styles):
    """Create Security Incidents tracking sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Secure Area Security Incidents\nDocument all incidents from last 12 months"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Incident ID": 15,
        "Date": 12,
        "Secure Area": 25,
        "Incident Type": 22,
        "Severity": 12,
        "Description": 40,
        "Response Time (min)": 18,
        "Resolution Status": 15,
        "Corrective Action": 40,
        "Notes": 35,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Unauthorised Access,Procedure Violation,Tailgating,Equipment Policy Violation,Recording Violation,Other"',
        allow_blank=True
    )
    dv_severity = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Resolved,In Progress,Escalated"', allow_blank=True)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_severity)
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=4))
        dv_severity.add(ws.cell(row=r, column=5))
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"


def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with metrics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "SECURE AREAS WORKING - SUMMARY DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Period:"
    ws["A3"].font = Font(name="Calibri", bold=True)
    ws["B3"].fill = styles["input_cell"]["fill"]
    ws["B3"].border = styles["border"]

    ws["A5"] = "Compliance Metrics"
    ws["A5"].font = Font(name="Calibri", size=12, bold=True)

    headers = ["Metric", "Total", f"{CHECK} Compliant", f"{WARNING} Partial", f"{XMARK} Non-Compliant", "N/A", "% Compliant"]
    row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    areas = [
        ("Secure Areas", "Secure Area Register", 11),
        ("Working Procedures", "Working Procedures", 5),
        ("Third-Party Access", "Third-Party Access", 8),
    ]

    row += 1
    start_data_row = row
    for label, sheet, status_col in areas:
        ws.cell(row=row, column=1, value=label)
        status_col_letter = get_column_letter(status_col)
        status_range = f"'{sheet}'!{status_col_letter}4:{status_col_letter}103"

        ws.cell(row=row, column=2, value=f'=COUNTA({status_range})')
        ws.cell(row=row, column=3, value=f'=COUNTIF({status_range},"{CHECK}*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({status_range},"{WARNING}*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({status_range},"{XMARK}*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({status_range},"N/A")')
        ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{row-1})")
        cell.font = Font(name="Calibri", bold=True)

    total_pct = ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
    total_pct.font = Font(name="Calibri", bold=True, color="0000FF", size=12)

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 18

    # Incident metrics
    row += 3
    ws[f"A{row}"] = "Incident Metrics"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)

    row += 1
    ws[f"A{row}"] = "Total Incidents (Last 12 Months):"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws[f"B{row}"] = "=COUNTA(Incidents!A4:A103)"

    row += 1
    ws[f"A{row}"] = "Critical/High Incidents:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws[f"B{row}"] = '=COUNTIF(Incidents!E4:E103,"Critical")+COUNTIF(Incidents!E4:E103,"High")'

    row += 1
    ws[f"A{row}"] = "Average Response Time (min):"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws[f"B{row}"] = "=IFERROR(AVERAGE(Incidents!G4:G103),0)"


def create_evidence_register(ws, styles):
    """Create Evidence Register sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID",
        "Evidence Type",
        "Description",
        "Related Sheet/Item",
        "File Name",
        "File Location",
        "Collection Date",
        "Collected By",
        "Retention Period",
        "Notes",
    ]
    widths = [15, 20, 40, 25, 30, 45, 16, 20, 18, 35]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Floor Plan,Access Log,Configuration Screenshot,Policy Document,Procedure Document,Photo,Audit Report,Training Record,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EV-{r-4:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "A5"


def create_approval_signoff(ws, styles):
    """Create Approval and Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Summary"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Overall Compliance Rate", "='Summary Dashboard'!G10"),
        ("Assessment Status", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Approval Workflow"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    approval_headers = ["Role", "Name", "Signature", "Date"]
    row += 3
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    roles = ["Assessor", "Security Operations Manager", "CISO", "Compliance Officer"]
    row += 1
    for role in roles:
        ws.cell(row=row, column=1, value=role).font = Font(name="Calibri", bold=True)
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15


# =============================================================================
# MAIN WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    styles = setup_styles()

    # Create all sheets
    ws = wb.create_sheet("Instructions & Legend", 0)
    create_instructions_sheet(ws, styles)

    ws = wb.create_sheet("Secure Area Register")
    create_secure_area_register(ws, styles)

    ws = wb.create_sheet("Working Procedures")
    create_working_procedures(ws, styles)

    ws = wb.create_sheet("Third-Party Access")
    create_third_party_access(ws, styles)

    ws = wb.create_sheet("Incidents")
    create_incidents_sheet(ws, styles)

    ws = wb.create_sheet("Summary Dashboard")
    create_summary_dashboard(ws, styles)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)

    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_signoff(ws, styles)

    return wb


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main entry point for workbook generation."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS Secure Areas Working Assessment (A.7.6.S1)")
        logger.info("=" * 70)

        wb = create_workbook()
        wb.save(OUTPUT_FILENAME)

        logger.info("%s SUCCESS: %s", CHECK, OUTPUT_FILENAME)
        logger.info("  %s 8 professional worksheets created", BULLET)
        logger.info("  %s Navy blue headers, yellow input cells", BULLET)
        logger.info("  %s 100 data rows per assessment sheet", BULLET)
        logger.info("  %s Automated compliance dashboard with formulas", BULLET)
        logger.info("  %s Data validations and freeze panes configured", BULLET)
        logger.info("  %s Evidence register with audit traceability", BULLET)
        logger.info("  %s 4-level approval workflow", BULLET)
        logger.info("=" * 70)

        return 0

    except Exception as e:
        logger.error("%s ERROR: Failed to generate workbook", XMARK)
        logger.error("Error: %s", str(e))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation with constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
