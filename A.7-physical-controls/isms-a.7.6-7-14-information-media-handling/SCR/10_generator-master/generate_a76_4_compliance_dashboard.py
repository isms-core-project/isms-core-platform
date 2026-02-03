#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.6-7-14.S4 - Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.7.6, A.7.7, A.7.14
Consolidated Dashboard: Secure Areas and Media Handling

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates a consolidated compliance dashboard that integrates data
from the individual S1, S2, and S3 assessments.

Reference: ISMS-IMP-A.7.6-7-14-S4 Specification

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    pip3 install openpyxl

Usage:
    python3 generate_a76_4_compliance_dashboard.py

Output:
    ISMS-IMP-A.7.6-7-14.S4_Compliance_Dashboard_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.6-7-14.S4"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.7.6-7-14"
CONTROL_NAME = "Secure Areas and Media Handling"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = '\u2705'
XMARK = '\u274C'
WARNING = '\u26A0'
BULLET = '\u2022'
ARROW = '\u2192'

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Dashboard Type", "Consolidated Compliance Dashboard"),
        ("Related Policies", "ISMS-POL-A.7.6-7-14"),
        ("Version", "1.0"),
        ("Generated Date", GENERATED_DATE),
        ("Generated By", ""),
        ("Organisation", ""),
        ("Reporting Period", ""),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "DASHBOARD PURPOSE"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    purpose = [
        "This dashboard consolidates compliance metrics from:",
        f"  {BULLET} A.7.6 - Secure Areas Working (S1 Assessment)",
        f"  {BULLET} A.7.7 - Clear Desk/Screen (S2 Assessment)",
        f"  {BULLET} A.7.14 - Equipment Disposal (S3 Assessment)",
        "",
        "Use this dashboard to:",
        f"  {BULLET} Track overall compliance score across all controls",
        f"  {BULLET} Monitor key performance indicators (KPIs)",
        f"  {BULLET} Identify and track gaps requiring remediation",
        f"  {BULLET} Prepare evidence for audit",
    ]

    row += 2
    for line in purpose:
        ws[f"A{row}"] = line
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50


def create_executive_dashboard(ws, styles):
    """Create Executive Dashboard sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "SECURE AREAS AND MEDIA HANDLING - EXECUTIVE DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "ISO/IEC 27001:2022 Controls A.7.6 / A.7.7 / A.7.14"
    ws["A2"].font = styles["subheader"]["font"]
    ws["A2"].fill = styles["subheader"]["fill"]
    ws["A2"].alignment = styles["subheader"]["alignment"]

    ws["A4"] = "Assessment Period:"
    ws["A4"].font = Font(name="Calibri", bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    # Control Compliance Summary
    ws["A6"] = "Control Compliance Summary"
    ws["A6"].font = Font(name="Calibri", size=12, bold=True)

    headers = ["Control", "Description", "Compliance Score", "Status", "Critical Gaps", "High Priority", "Medium Priority"]
    row = 7
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    controls = [
        ("A.7.6", "Working in Secure Areas"),
        ("A.7.7", "Clear Desk and Clear Screen"),
        ("A.7.14", "Secure Disposal of Equipment"),
    ]

    row = 8
    dv_status = DataValidation(type="list", formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant"', allow_blank=True)
    ws.add_data_validation(dv_status)

    for control, desc in controls:
        ws.cell(row=row, column=1, value=control).border = styles["border"]
        ws.cell(row=row, column=2, value=desc).border = styles["border"]
        for c in range(3, 8):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_status.add(ws.cell(row=row, column=4))
        row += 1

    # Overall row
    ws.cell(row=row, column=1, value="OVERALL").font = Font(name="Calibri", bold=True)
    ws.cell(row=row, column=1).border = styles["border"]
    ws.cell(row=row, column=2, value="Secure Areas and Media Handling").border = styles["border"]
    for c in range(3, 8):
        cell = ws.cell(row=row, column=c)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = styles["border"]
        cell.font = Font(name="Calibri", bold=True)
    dv_status.add(ws.cell(row=row, column=4))

    # Key Metrics
    row += 3
    ws[f"A{row}"] = "Key Performance Indicators"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)

    metrics_headers = ["Metric", "Current Value", "Target", "Status", "Trend"]
    row += 1
    for col_idx, header in enumerate(metrics_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    metrics = [
        ("Secure Area Compliance", "", "100%"),
        ("Third-Party Escort Compliance", "", "100%"),
        ("Clear Desk Audit Pass Rate", "", ">95%"),
        ("Screen Lock Compliance", "", "100%"),
        ("Equipment Disposal with Certificate", "", "100%"),
        ("Secure Area Incidents (YTD)", "", "0"),
    ]

    dv_trend = DataValidation(type="list", formula1=f'"{ARROW} Improving,= Stable,{ARROW} Declining"', allow_blank=True)
    ws.add_data_validation(dv_trend)

    row += 1
    for metric, current, target in metrics:
        ws.cell(row=row, column=1, value=metric).border = styles["border"]
        for c in range(2, 6):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        ws.cell(row=row, column=3, value=target).border = styles["border"]
        dv_status.add(ws.cell(row=row, column=4))
        dv_trend.add(ws.cell(row=row, column=5))
        row += 1

    # Set column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15


def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "GAP ANALYSIS AND REMEDIATION TRACKING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    columns = {
        "Gap ID": 12,
        "Control": 12,
        "Finding": 45,
        "Severity": 12,
        "Owner": 20,
        "Target Date": 15,
        "Status": 15,
        "Progress Notes": 40,
        "Evidence": 30,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_control = DataValidation(type="list", formula1='"A.7.6,A.7.7,A.7.14"', allow_blank=True)
    dv_severity = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Open,In Progress,Closed,On Hold"', allow_blank=True)
    ws.add_data_validation(dv_control)
    ws.add_data_validation(dv_severity)
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"GAP-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_control.add(ws.cell(row=r, column=2))
        dv_severity.add(ws.cell(row=r, column=4))
        dv_status.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "A4"


def create_kpis_metrics(ws, styles):
    """Create KPIs & Metrics sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "KEY PERFORMANCE INDICATORS AND METRICS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    columns = {
        "KPI ID": 10,
        "Control": 12,
        "KPI Description": 40,
        "Measurement": 25,
        "Current Value": 15,
        "Target": 15,
        "Status": 15,
        "Trend (vs Last Period)": 20,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate KPIs
    kpis = [
        ("KPI-01", "A.7.6", "Secure area access compliance", "% of accesses compliant", "", "100%"),
        ("KPI-02", "A.7.6", "Third-party escort compliance", "% visitors escorted", "", "100%"),
        ("KPI-03", "A.7.6", "Secure area incident count", "Count per month", "", "0"),
        ("KPI-04", "A.7.6", "Access review completion", "% completed on time", "", "100%"),
        ("KPI-05", "A.7.7", "Clear desk audit frequency", "Audits per month", "", ">=1"),
        ("KPI-06", "A.7.7", "Clear desk audit pass rate", "% passing audit", "", ">95%"),
        ("KPI-07", "A.7.7", "Screen lock compliance", "% devices compliant", "", "100%"),
        ("KPI-08", "A.7.7", "Workspace with lockable storage", "% workspaces", "", "100%"),
        ("KPI-09", "A.7.14", "Disposal with certificate rate", "% with certificate", "", "100%"),
        ("KPI-10", "A.7.14", "Disposal policy compliance", "% compliant disposals", "", "100%"),
        ("KPI-11", "A.7.14", "Active vendor contracts", "Count with valid contract", "", ">1"),
        ("KPI-12", "A.7.14", "Average disposal response time", "Days from request", "", "<7"),
    ]

    dv_status = DataValidation(type="list", formula1=f'"{CHECK} Met,{WARNING} Near Target,{XMARK} Not Met"', allow_blank=True)
    dv_trend = DataValidation(type="list", formula1='"Improving,Stable,Declining,New"', allow_blank=True)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_trend)

    row = 4
    for kpi_id, control, desc, measure, current, target in kpis:
        ws.cell(row=row, column=1, value=kpi_id).border = styles["border"]
        ws.cell(row=row, column=2, value=control).border = styles["border"]
        ws.cell(row=row, column=3, value=desc).border = styles["border"]
        ws.cell(row=row, column=4, value=measure).border = styles["border"]
        ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=5).border = styles["border"]
        ws.cell(row=row, column=6, value=target).border = styles["border"]
        ws.cell(row=row, column=7).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=7).border = styles["border"]
        ws.cell(row=row, column=8).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=8).border = styles["border"]
        dv_status.add(ws.cell(row=row, column=7))
        dv_trend.add(ws.cell(row=row, column=8))
        row += 1

    # Additional blank rows
    for r in range(row, row + 38):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_status.add(ws.cell(row=r, column=7))
        dv_trend.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"


def create_evidence_register(ws, styles):
    """Create Evidence Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "CONSOLIDATED EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "Consolidate key evidence from S1, S2, S3 assessments for audit readiness."
    ws["A2"].font = Font(name="Calibri", italic=True)

    columns = {
        "Evidence ID": 15,
        "Control": 12,
        "Evidence Type": 22,
        "Description": 45,
        "Source Assessment": 20,
        "Date Collected": 15,
        "Location/Link": 45,
        "Status": 18,
    }

    row = 4
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_control = DataValidation(type="list", formula1='"A.7.6,A.7.7,A.7.14,Multiple"', allow_blank=True)
    dv_source = DataValidation(type="list", formula1='"S1 - Secure Areas,S2 - Clear Desk/Screen,S3 - Equipment Disposal,Dashboard,External"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Collected,Pending,Requires Update,Missing"', allow_blank=True)
    ws.add_data_validation(dv_control)
    ws.add_data_validation(dv_source)
    ws.add_data_validation(dv_status)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EV-CONS-{r-4:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_control.add(ws.cell(row=r, column=2))
        dv_source.add(ws.cell(row=r, column=5))
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


def create_approval_signoff(ws, styles):
    """Create Approval and Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "DASHBOARD APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A3"] = "Dashboard Summary"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    summary_fields = [
        ("Dashboard Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Reporting Period", ""),
        ("A.7.6 Compliance Score", ""),
        ("A.7.7 Compliance Score", ""),
        ("A.7.14 Compliance Score", ""),
        ("Overall Compliance Score", ""),
        ("Open Critical Gaps", ""),
        ("Dashboard Status", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Approval Workflow"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    approval_headers = ["Role", "Name", "Signature", "Date", "Comments"]
    row += 3
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    roles = [
        "Facilities Manager",
        "IT Operations Manager",
        "Security Operations Manager",
        "CISO",
        "Compliance Officer",
    ]
    row += 1
    for role in roles:
        ws.cell(row=row, column=1, value=role).font = Font(name="Calibri", bold=True)
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 35


# =============================================================================
# MAIN WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    styles = setup_styles()

    ws = wb.create_sheet("Instructions & Legend", 0)
    create_instructions_sheet(ws, styles)

    ws = wb.create_sheet("Executive Dashboard")
    create_executive_dashboard(ws, styles)

    ws = wb.create_sheet("Gap Analysis")
    create_gap_analysis(ws, styles)

    ws = wb.create_sheet("KPIs & Metrics")
    create_kpis_metrics(ws, styles)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)

    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_signoff(ws, styles)

    return wb


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main entry point for workbook generation."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS Secure Areas and Media Handling Dashboard (A.7.6-7-14.S4)")
        logger.info("=" * 70)

        wb = create_workbook()
        wb.save(OUTPUT_FILENAME)

        logger.info("%s SUCCESS: %s", CHECK, OUTPUT_FILENAME)
        logger.info("  %s 6 professional worksheets created", BULLET)
        logger.info("  %s Executive dashboard with control summary", BULLET)
        logger.info("  %s Gap analysis with remediation tracking", BULLET)
        logger.info("  %s 12 pre-populated KPIs across all controls", BULLET)
        logger.info("  %s Consolidated evidence register", BULLET)
        logger.info("  %s 5-level approval workflow", BULLET)
        logger.info("=" * 70)

        return 0

    except Exception as e:
        logger.error("%s ERROR: Failed to generate workbook", XMARK)
        logger.error("Error: %s", str(e))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation with constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
