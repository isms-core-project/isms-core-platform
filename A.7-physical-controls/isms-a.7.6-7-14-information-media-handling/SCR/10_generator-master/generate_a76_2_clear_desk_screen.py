#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.7.S2 - Clear Desk and Clear Screen Compliance Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.7: Clear Desk and Clear Screen
Assessment Domain 2 of 4: Workspace Information Protection

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates an assessment workbook for documenting clear desk/screen
requirements, screen lock configurations, audit results, and workspace compliance.

Reference: ISMS-IMP-A.7.6-7-14-S2 Specification

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    pip3 install openpyxl

Usage:
    python3 generate_a76_2_clear_desk_screen.py

Output:
    ISMS-IMP-A.7.7.S2_Clear_Desk_Screen_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.7.S2"
WORKBOOK_NAME = "Clear Desk Screen Compliance"
CONTROL_ID = "A.7.7"
CONTROL_NAME = "Clear Desk and Clear Screen"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = '\u2705'
XMARK = '\u274C'
WARNING = '\u26A0'
BULLET = '\u2022'

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Clear Desk and Clear Screen Compliance"),
        ("Related Policy", "ISMS-POL-A.7.6-7-14"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "1. Document requirements by classification in Sheet 2 (Requirements Matrix).",
        "2. Verify screen lock configurations in Sheet 3 (Screen Lock Configuration).",
        "3. Track audit results in Sheet 4 (Audit Results).",
        "4. Assess workspace infrastructure in Sheet 5 (Workspace Assessment).",
        "5. Review Summary Dashboard (Sheet 6) for compliance metrics.",
        "6. Document evidence in Sheet 7 (Evidence Register).",
        "7. Obtain approvals in Sheet 8 (Approval Sign-Off).",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50


def create_requirements_matrix(ws, styles):
    """Create Requirements Matrix sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Clear Desk/Screen Requirements Matrix\nRequirements by Information Classification Level"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Classification": 18,
        "Clear Desk - Work Hours": 25,
        "Clear Desk - End of Day": 25,
        "Storage Requirement": 25,
        "Screen Lock Timeout (min)": 22,
        "Privacy Screen Required": 22,
        "Implementation Status": 20,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate classification levels
    classifications = [
        ("CONFIDENTIAL", "Locked when unattended", "Mandatory locked storage", "Lockable drawer/cabinet", 5, "Required in open areas"),
        ("INTERNAL", "Store when not in use", "Locked at end of day", "Lockable drawer/cabinet", 10, "Recommended"),
        ("PUBLIC", "Best practice", "Not mandatory", "Open storage acceptable", 15, "Not required"),
    ]

    dv_impl = DataValidation(type="list", formula1='"Implemented,Partial,Not Implemented"', allow_blank=True)
    dv_privacy = DataValidation(type="list", formula1='"Required,Recommended,Not Required"', allow_blank=True)
    ws.add_data_validation(dv_impl)
    ws.add_data_validation(dv_privacy)

    row = 4
    for clf, work_hrs, end_day, storage, timeout, privacy in classifications:
        ws.cell(row=row, column=1, value=clf).border = styles["border"]
        ws.cell(row=row, column=2, value=work_hrs).border = styles["border"]
        ws.cell(row=row, column=3, value=end_day).border = styles["border"]
        ws.cell(row=row, column=4, value=storage).border = styles["border"]
        ws.cell(row=row, column=5, value=timeout).border = styles["border"]
        ws.cell(row=row, column=6, value=privacy).border = styles["border"]
        for c in range(7, 9):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_impl.add(ws.cell(row=row, column=7))
        row += 1

    ws.freeze_panes = "A4"


def create_screen_lock_config(ws, styles):
    """Create Screen Lock Configuration sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Screen Lock Configuration Assessment\nVerify screen lock timeout settings across device types"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Device Type": 20,
        "Policy/GPO Name": 30,
        "Configured Timeout (min)": 22,
        "Required Timeout (min)": 22,
        "Compliant": 12,
        "Enforcement Method": 20,
        "Device Count": 15,
        "Last Verified": 15,
        "Evidence": 30,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate device types
    device_types = [
        "Windows Workstations",
        "MacOS Devices",
        "Mobile Devices (iOS)",
        "Mobile Devices (Android)",
        "Linux Workstations",
    ]

    dv_enforcement = DataValidation(type="list", formula1='"Group Policy,MDM (Intune),MDM (JAMF),MDM (Other),Manual,Not Enforced"', allow_blank=True)
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_enforcement)
    ws.add_data_validation(dv_yesno)

    row = 4
    for device in device_types:
        ws.cell(row=row, column=1, value=device).border = styles["border"]
        for c in range(2, 11):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        # Compliance formula
        ws.cell(row=row, column=5, value=f'=IF(C{row}<=D{row},"Yes","No")')
        dv_enforcement.add(ws.cell(row=row, column=6))
        row += 1

    # Additional blank rows
    for r in range(row, row + 45):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        ws.cell(row=r, column=5, value=f'=IF(C{r}<=D{r},"Yes","No")')
        dv_enforcement.add(ws.cell(row=r, column=6))

    ws.freeze_panes = "A4"


def create_audit_results(ws, styles):
    """Create Audit Results tracking sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Clear Desk Audit Results\nTrack monthly audit results and compliance trends"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Audit Date": 15,
        "Audit Type": 20,
        "Location/Area": 25,
        "Workstations Audited": 20,
        "Compliant": 12,
        "Non-Compliant": 15,
        "Compliance Rate": 15,
        "Common Issues": 35,
        "Follow-Up Actions": 35,
        "Auditor": 20,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(type="list", formula1='"Scheduled Monthly,Random Spot Check,Post-Incident,Annual Comprehensive"', allow_blank=True)
    ws.add_data_validation(dv_type)

    for r in range(4, 104):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_type.add(ws.cell(row=r, column=2))
        # Compliance rate formula
        ws.cell(row=r, column=7, value=f'=IF(D{r}=0,"",ROUND(E{r}/D{r}*100,1)&"%")')

    ws.freeze_panes = "A4"


def create_workspace_assessment(ws, styles):
    """Create Workspace Assessment sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "Workspace Infrastructure Assessment\nAssess compliance infrastructure by workspace area"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Area ID": 12,
        "Location": 30,
        "Workspace Type": 20,
        "Workstation Count": 18,
        "Lockable Storage": 18,
        "Shredder Access": 18,
        "Confidential Bins": 18,
        "Privacy Screens": 18,
        "Avg Audit Compliance": 20,
        "Status": 15,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(type="list", formula1='"Open Plan Office,Private Office,Meeting Room,Shared Space,Home Office,Reception"', allow_blank=True)
    dv_storage = DataValidation(type="list", formula1='"Yes - All,Yes - Partial,No"', allow_blank=True)
    dv_shredder = DataValidation(type="list", formula1='"Yes - On Floor,Yes - Central,No"', allow_blank=True)
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_privacy = DataValidation(type="list", formula1='"Yes - All,Yes - Partial,No,N/A"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"', allow_blank=True)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_storage)
    ws.add_data_validation(dv_shredder)
    ws.add_data_validation(dv_yesno)
    ws.add_data_validation(dv_privacy)
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_type.add(ws.cell(row=r, column=3))
        dv_storage.add(ws.cell(row=r, column=5))
        dv_shredder.add(ws.cell(row=r, column=6))
        dv_yesno.add(ws.cell(row=r, column=7))
        dv_privacy.add(ws.cell(row=r, column=8))
        dv_status.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "A4"


def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with metrics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "CLEAR DESK/SCREEN - SUMMARY DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Period:"
    ws["A3"].font = Font(name="Calibri", bold=True)
    ws["B3"].fill = styles["input_cell"]["fill"]
    ws["B3"].border = styles["border"]

    ws["A5"] = "Compliance Metrics"
    ws["A5"].font = Font(name="Calibri", size=12, bold=True)

    # Screen Lock Compliance
    ws["A7"] = "Screen Lock Configuration Compliance"
    ws["A7"].font = Font(name="Calibri", size=11, bold=True)

    ws["A8"] = "Total Device Types:"
    ws["B8"] = "=COUNTA('Screen Lock Configuration'!A4:A53)"
    ws["A9"] = "Compliant:"
    ws["B9"] = '=COUNTIF(\'Screen Lock Configuration\'!E4:E53,"Yes")'
    ws["A10"] = "Non-Compliant:"
    ws["B10"] = '=COUNTIF(\'Screen Lock Configuration\'!E4:E53,"No")'
    ws["A11"] = "Compliance Rate:"
    ws["B11"] = '=IF(B8=0,"0%",ROUND(B9/B8*100,1)&"%")'
    ws["B11"].font = Font(name="Calibri", bold=True, color="0000FF")

    # Audit Results Summary
    ws["A13"] = "Clear Desk Audit Summary"
    ws["A13"].font = Font(name="Calibri", size=11, bold=True)

    ws["A14"] = "Total Audits (Last 12 Months):"
    ws["B14"] = "=COUNTA('Audit Results'!A4:A103)"
    ws["A15"] = "Average Compliance Rate:"
    ws["B15"] = "=IFERROR(AVERAGE('Audit Results'!E4:E103)/AVERAGE('Audit Results'!D4:D103)*100,0)&\"%\""
    ws["B15"].font = Font(name="Calibri", bold=True, color="0000FF")

    # Workspace Assessment
    ws["A17"] = "Workspace Infrastructure"
    ws["A17"].font = Font(name="Calibri", size=11, bold=True)

    ws["A18"] = "Total Workspaces Assessed:"
    ws["B18"] = "=COUNTA('Workspace Assessment'!A4:A103)"
    ws["A19"] = "With Full Lockable Storage:"
    ws["B19"] = '=COUNTIF(\'Workspace Assessment\'!E4:E103,"Yes - All")'
    ws["A20"] = "With Privacy Screens:"
    ws["B20"] = '=COUNTIF(\'Workspace Assessment\'!H4:H103,"Yes*")'

    for col in ["A", "B"]:
        ws.column_dimensions[col].width = 35


def create_evidence_register(ws, styles):
    """Create Evidence Register sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment."
    ws["A2"].font = Font(name="Calibri", italic=True)

    headers = [
        "Evidence ID", "Evidence Type", "Description", "Related Sheet/Item",
        "File Name", "File Location", "Collection Date", "Collected By",
        "Retention Period", "Notes",
    ]
    widths = [15, 20, 40, 25, 30, 45, 16, 20, 18, 35]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Policy Document,GPO/MDM Config,Audit Checklist,Audit Report,Training Record,Photo,Screenshot,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EV-{r-4:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "A5"


def create_approval_signoff(ws, styles):
    """Create Approval and Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Summary"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Screen Lock Compliance", "='Summary Dashboard'!B11"),
        ("Assessment Status", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Approval Workflow"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    approval_headers = ["Role", "Name", "Signature", "Date"]
    row += 3
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    roles = ["Assessor", "Facilities Manager", "CISO", "Compliance Officer"]
    row += 1
    for role in roles:
        ws.cell(row=row, column=1, value=role).font = Font(name="Calibri", bold=True)
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15


# =============================================================================
# MAIN WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    styles = setup_styles()

    ws = wb.create_sheet("Instructions & Legend", 0)
    create_instructions_sheet(ws, styles)

    ws = wb.create_sheet("Requirements Matrix")
    create_requirements_matrix(ws, styles)

    ws = wb.create_sheet("Screen Lock Configuration")
    create_screen_lock_config(ws, styles)

    ws = wb.create_sheet("Audit Results")
    create_audit_results(ws, styles)

    ws = wb.create_sheet("Workspace Assessment")
    create_workspace_assessment(ws, styles)

    ws = wb.create_sheet("Summary Dashboard")
    create_summary_dashboard(ws, styles)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)

    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_signoff(ws, styles)

    return wb


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main entry point for workbook generation."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS Clear Desk/Screen Compliance Assessment (A.7.7.S2)")
        logger.info("=" * 70)

        wb = create_workbook()
        wb.save(OUTPUT_FILENAME)

        logger.info("%s SUCCESS: %s", CHECK, OUTPUT_FILENAME)
        logger.info("  %s 8 professional worksheets created", BULLET)
        logger.info("  %s Requirements matrix by classification", BULLET)
        logger.info("  %s Screen lock configuration tracking", BULLET)
        logger.info("  %s Audit results with compliance formulas", BULLET)
        logger.info("  %s Workspace infrastructure assessment", BULLET)
        logger.info("  %s Automated summary dashboard", BULLET)
        logger.info("=" * 70)

        return 0

    except Exception as e:
        logger.error("%s ERROR: Failed to generate workbook", XMARK)
        logger.error("Error: %s", str(e))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation with constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
