#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.1.3 - Secure Areas Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.3: Securing Offices, Rooms and Facilities
Assessment Domain 3 of 4: Office and Facility Security

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific facility security requirements, office layouts,
and sensitive area classifications.

Key customization areas:
1. Office/room classifications (match your facility designations)
2. Security requirements by area type (aligned with your policy)
3. Server room and datacenter controls (specific to your infrastructure)
4. Shared facility considerations (based on your building arrangements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.7.4 Physical Security Monitoring Framework

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (Python standard library)

Usage:
    python3 generate_a71_3_secure_areas.py

Output:
    ISMS-IMP-A.7.1.3_Secure_Areas_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.1.3"
WORKBOOK_NAME = "Secure Areas Assessment"
CONTROL_ID = "A.7.3"
CONTROL_NAME = "Securing Offices, Rooms and Facilities"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # Green checkmark
XMARK = '\u274C'      # Red X
WARNING = '\u26A0'    # Warning sign
HOURGLASS = '\u23F3'  # Hourglass
BULLET = '\u2022'     # Bullet point
ARROW = '\u2192'      # Right arrow

# ============================================================================
# STYLE DEFINITIONS
# ============================================================================

def setup_styles():
    """Define all cell styles matching framework pattern."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# ============================================================================
# INSTRUCTIONS SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.7.1.3 - Secure Areas Assessment\n"
        "ISO/IEC 27001:2022 - Control A.7.3: Securing Offices, Rooms and Facilities"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.7.1.3"),
        ("Assessment Area", "Office, Room and Facility Security"),
        ("Related Policy", "ISMS-POL-A.7.1-2-3"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Annual or after facility changes"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "1. Complete each worksheet tab for applicable secure areas.",
        "2. Document all offices handling sensitive information.",
        "3. Assess server rooms and datacenters against requirements.",
        "4. Review meeting room security procedures.",
        "5. Evaluate shared facility security arrangements.",
        "6. Use dropdown menus for standardised status entries.",
        "7. Fill in yellow-highlighted cells with your facility information.",
        "8. Summary Dashboard auto-calculates compliance metrics.",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws[f"A{row+1}"] = "STATUS LEGEND"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    legend = [
        ("Symbol", "Status", "Description"),
        (f"{CHECK}", "Compliant", "Meets all secure area requirements"),
        (f"{WARNING}", "Partial", "Partial compliance, requires improvement"),
        (f"{XMARK}", "Non-Compliant", "Does not meet requirements, immediate action needed"),
        ("N/A", "Not Applicable", "Not required for this area type"),
    ]

    row += 2
    for c, h in enumerate(legend[0], start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for sym, status, desc in legend[1:]:
        ws.cell(row=row, column=1, value=sym).border = styles["border"]
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if "Compliant" in status and "Non" not in status:
            s.fill = styles["status_compliant"]["fill"]
        elif "Partial" in status:
            s.fill = styles["status_partial"]["fill"]
        elif "Non-Compliant" in status:
            s.fill = styles["status_noncompliant"]["fill"]
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 50

# ============================================================================
# OFFICE SECURITY SHEET
# ============================================================================

def create_office_security_sheet(ws, styles):
    """Create Office Security assessment sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Office Security Assessment\nPolicy Requirement: Offices locked when unoccupied, clean desk enforced, screens protected"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A3:J3")
    ws["A3"] = "Are office security controls implemented including locking, clean desk policy, and screen privacy?"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    ws["A4"] = "Response:"
    ws["A4"].font = Font(name="Calibri", bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    dv_response = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=False)
    ws.add_data_validation(dv_response)
    dv_response.add(ws["B4"])

    columns = {
        "Office/Area ID": 15,
        "Location": 25,
        "Security Zone": 18,
        "Classification": 18,
        "Lock When Unoccupied": 18,
        "Clean Desk Enforced": 18,
        "Screen Privacy": 15,
        "Secure Storage": 15,
        "Status": 18,
        "Notes": 35,
    }

    row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(7, 107):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # Security Zone dropdown (column C)
    dv_zone = DataValidation(
        type="list",
        formula1='"Controlled Zone,Restricted Zone,High-Security Zone"',
        allow_blank=False
    )
    ws.add_data_validation(dv_zone)
    for r in range(7, 107):
        dv_zone.add(ws.cell(row=r, column=3))

    # Classification dropdown (column D)
    dv_class = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Highly Confidential"',
        allow_blank=False
    )
    ws.add_data_validation(dv_class)
    for r in range(7, 107):
        dv_class.add(ws.cell(row=r, column=4))

    # Yes/No dropdowns
    dv_yesno = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=False)
    ws.add_data_validation(dv_yesno)
    for r in range(7, 107):
        dv_yesno.add(ws.cell(row=r, column=5))
        dv_yesno.add(ws.cell(row=r, column=6))
        dv_yesno.add(ws.cell(row=r, column=7))
        dv_yesno.add(ws.cell(row=r, column=8))

    # Status dropdown (column I)
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)
    for r in range(7, 107):
        dv_status.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "A7"

# ============================================================================
# SERVER ROOMS SHEET
# ============================================================================

def create_server_rooms_sheet(ws, styles):
    """Create Server Rooms and Datacenters assessment sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "Server Rooms and Datacenters\nPolicy Requirement: MFA access, no windows, reinforced construction, environmental monitoring"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A3:K3")
    ws["A3"] = "Are server rooms and datacenters secured with multi-factor access, CCTV, and environmental controls?"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    ws["A4"] = "Response:"
    ws["A4"].font = Font(name="Calibri", bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    dv_response = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=False)
    ws.add_data_validation(dv_response)
    dv_response.add(ws["B4"])

    columns = {
        "Room ID": 12,
        "Location": 25,
        "Room Type": 18,
        "MFA Access": 12,
        "No Windows": 12,
        "Reinforced Walls": 15,
        "CCTV Coverage": 15,
        "Access Logging": 15,
        "Env. Monitoring": 15,
        "Status": 18,
        "Notes": 35,
    }

    row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(7, 107):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # Room Type dropdown (column C)
    dv_room = DataValidation(
        type="list",
        formula1='"Server Room,Datacenter,Network Closet,Comms Room,UPS Room,Generator Room"',
        allow_blank=False
    )
    ws.add_data_validation(dv_room)
    for r in range(7, 107):
        dv_room.add(ws.cell(row=r, column=3))

    # Yes/No dropdowns
    dv_yesno = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=False)
    ws.add_data_validation(dv_yesno)
    for r in range(7, 107):
        dv_yesno.add(ws.cell(row=r, column=4))
        dv_yesno.add(ws.cell(row=r, column=5))
        dv_yesno.add(ws.cell(row=r, column=6))
        dv_yesno.add(ws.cell(row=r, column=7))
        dv_yesno.add(ws.cell(row=r, column=8))
        dv_yesno.add(ws.cell(row=r, column=9))

    # Status dropdown (column J)
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)
    for r in range(7, 107):
        dv_status.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "A7"

# ============================================================================
# MEETING ROOMS SHEET
# ============================================================================

def create_meeting_rooms_sheet(ws, styles):
    """Create Meeting Rooms security assessment sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Meeting Room Security\nPolicy Requirement: Recording device checks, whiteboard clearing, equipment security"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A3:I3")
    ws["A3"] = "Are meeting room security procedures implemented for sensitive discussions?"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    ws["A4"] = "Response:"
    ws["A4"].font = Font(name="Calibri", bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    dv_response = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=False)
    ws.add_data_validation(dv_response)
    dv_response.add(ws["B4"])

    columns = {
        "Room ID": 12,
        "Location": 25,
        "Room Classification": 20,
        "Recording Check": 15,
        "Whiteboard Clear": 15,
        "Document Clear": 15,
        "VC Equipment Secured": 18,
        "Status": 18,
        "Notes": 35,
    }

    row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(7, 107):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # Room Classification dropdown (column C)
    dv_class = DataValidation(
        type="list",
        formula1='"Standard Meeting Room,Boardroom,Secure Meeting Room,Interview Room"',
        allow_blank=False
    )
    ws.add_data_validation(dv_class)
    for r in range(7, 107):
        dv_class.add(ws.cell(row=r, column=3))

    # Yes/No dropdowns
    dv_yesno = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=False)
    ws.add_data_validation(dv_yesno)
    for r in range(7, 107):
        dv_yesno.add(ws.cell(row=r, column=4))
        dv_yesno.add(ws.cell(row=r, column=5))
        dv_yesno.add(ws.cell(row=r, column=6))
        dv_yesno.add(ws.cell(row=r, column=7))

    # Status dropdown (column H)
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)
    for r in range(7, 107):
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A7"

# ============================================================================
# SHARED FACILITIES SHEET
# ============================================================================

def create_shared_facilities_sheet(ws, styles):
    """Create Shared Facilities assessment sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Shared Facility Security\nPolicy Requirement: Clear perimeter definition, building management access control"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A3:I3")
    ws["A3"] = "For shared buildings, is the organisation's perimeter clearly defined with controlled access?"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    ws["A4"] = "Response:"
    ws["A4"].font = Font(name="Calibri", bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    dv_response = DataValidation(type="list", formula1='"Yes,No,Partial,N/A - Sole Occupant"', allow_blank=False)
    ws.add_data_validation(dv_response)
    dv_response.add(ws["B4"])

    columns = {
        "Facility/Building": 25,
        "Sharing Arrangement": 20,
        "Perimeter Defined": 15,
        "Own Access Control": 18,
        "Building Mgmt Access": 20,
        "Shared Infrastructure": 20,
        "Key Management": 18,
        "Status": 18,
        "Notes": 35,
    }

    row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(7, 107):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # Sharing Arrangement dropdown (column B)
    dv_sharing = DataValidation(
        type="list",
        formula1='"Multi-Tenant Building,Co-Working Space,Shared Floor,Colocation Facility,Sole Occupant"',
        allow_blank=False
    )
    ws.add_data_validation(dv_sharing)
    for r in range(7, 107):
        dv_sharing.add(ws.cell(row=r, column=2))

    # Yes/No dropdowns
    dv_yesno = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=False)
    ws.add_data_validation(dv_yesno)
    for r in range(7, 107):
        dv_yesno.add(ws.cell(row=r, column=3))
        dv_yesno.add(ws.cell(row=r, column=4))

    # Building Mgmt Access dropdown (column E)
    dv_bldg = DataValidation(
        type="list",
        formula1='"Controlled and Logged,Controlled Only,Uncontrolled,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_bldg)
    for r in range(7, 107):
        dv_bldg.add(ws.cell(row=r, column=5))

    # Status dropdown (column H)
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)
    for r in range(7, 107):
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A7"

# ============================================================================
# SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with metrics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "SECURE AREAS SUMMARY DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Period:"
    ws["A3"].font = Font(name="Calibri", bold=True)
    ws["B3"].fill = styles["input_cell"]["fill"]
    ws["B3"].border = styles["border"]

    ws["A5"] = "Compliance Metrics"
    ws["A5"].font = Font(name="Calibri", size=12, bold=True)

    headers = ["Assessment Area", "Count", f"{CHECK} Compliant", f"{WARNING} Partial", f"{XMARK} Non-Compliant", "N/A", "% Compliant"]
    row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    areas = [
        ("Office Security", "Office Security", 9),
        ("Server Rooms", "Server Rooms", 10),
        ("Meeting Rooms", "Meeting Rooms", 8),
        ("Shared Facilities", "Shared Facilities", 8),
    ]

    row += 1
    start_data_row = row
    for label, sheet, status_col in areas:
        ws.cell(row=row, column=1, value=label)

        status_col_letter = get_column_letter(status_col)
        status_range = f"'{sheet}'!{status_col_letter}7:{status_col_letter}106"

        ws.cell(row=row, column=2, value=f'=COUNTA({status_range})')
        ws.cell(row=row, column=3, value=f'=COUNTIF({status_range},"{CHECK}*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({status_range},"{WARNING}*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({status_range},"{XMARK}*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({status_range},"N/A")')
        ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{row-1})")
        cell.font = Font(name="Calibri", bold=True)

    total_pct = ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
    total_pct.font = Font(name="Calibri", bold=True, color="0000FF", size=12)

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 18

# ============================================================================
# EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID",
        "Assessment Area",
        "Evidence Type",
        "Description",
        "Location/Path",
        "Date Collected",
        "Collected By",
        "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Floor plan,Photograph,Access log,Configuration screenshot,Inspection report,Policy document,Other"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_type)

    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_ver)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EV-{r-4:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"

# ============================================================================
# APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create Approval and Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Summary"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    summary_fields = [
        ("Assessment Document", "ISMS-IMP-A.7.1.3 - Secure Areas Assessment"),
        ("Assessment Period", ""),
        ("Overall Compliance Rate", "='Summary Dashboard'!G11"),
        ("Assessment Status", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Approval Workflow"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    approval_headers = ["Role", "Name", "Signature", "Date"]
    row += 3
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    roles = ["Assessor", "Facilities Manager", "Security Manager", "CISO Approver"]
    row += 1
    for role in roles:
        ws.cell(row=row, column=1, value=role).font = Font(name="Calibri", bold=True)
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15

# ============================================================================
# MAIN WORKBOOK GENERATION
# ============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    styles = setup_styles()

    # Create all sheets
    ws = wb.create_sheet("Instructions & Legend", 0)
    create_instructions_sheet(ws, styles)

    ws = wb.create_sheet("Office Security")
    create_office_security_sheet(ws, styles)

    ws = wb.create_sheet("Server Rooms")
    create_server_rooms_sheet(ws, styles)

    ws = wb.create_sheet("Meeting Rooms")
    create_meeting_rooms_sheet(ws, styles)

    ws = wb.create_sheet("Shared Facilities")
    create_shared_facilities_sheet(ws, styles)

    ws = wb.create_sheet("Summary Dashboard")
    create_summary_dashboard(ws, styles)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)

    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_signoff(ws, styles)

    return wb

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main entry point for workbook generation."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS Secure Areas Assessment (A.7.3)")
        logger.info("=" * 70)

        wb = create_workbook()
        filename = f"ISMS-IMP-A.7.1.3_Secure_Areas_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)

        logger.info("%s SUCCESS: %s", CHECK, filename)
        logger.info("  %s 8 professional worksheets created", BULLET)
        logger.info("  %s Navy headers, yellow input cells styling", BULLET)
        logger.info("  %s 100 data rows per assessment sheet", BULLET)
        logger.info("  %s Automated compliance dashboard with formulas", BULLET)
        logger.info("  %s Data validations and freeze panes configured", BULLET)
        logger.info("  %s Evidence register with audit traceability", BULLET)
        logger.info("  %s 4-level approval workflow", BULLET)
        logger.info("=" * 70)

        return 0

    except Exception as e:
        logger.error("%s ERROR: Failed to generate workbook", XMARK)
        logger.error("Error: %s", str(e))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
