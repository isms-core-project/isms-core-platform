#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.9.S2 - Off-Premises Asset Security Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.9: Security of Assets Off-Premises
Assessment Domain 2 of 3: Off-Premises Asset Security

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific off-premises asset requirements, MDM platforms,
and assessment criteria.

Key customization areas:
1. Equipment categories (match your device types and MDM classifications)
2. Protection requirements (adapt to your security policies)
3. Remote working scenarios (based on your work arrangements)
4. Compliance thresholds (aligned with your policy requirements)

Reference Pattern: Based on ISMS-A.7.4 Physical Infrastructure Framework

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Usage:
    python3 generate_a78_2_offpremises_assets.py

Output:
    ISMS-IMP-A.7.9.S2_Off_Premises_Assets_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.9.S2"
WORKBOOK_NAME = "Off-Premises Asset Security"
CONTROL_ID = "A.7.9"
CONTROL_NAME = "Security of Assets Off-Premises"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_').replace('-', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = '\u2705'
XMARK = '\u274C'
WARNING = '\u26A0'
BULLET = '\u2022'

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.7.9.S2 - Off-Premises Asset Security Assessment\n"
        "ISO/IEC 27001:2022 - Control A.7.9: Security of Assets Off-Premises"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.7.9.S2"),
        ("Assessment Area", "Off-Premises Assets - Authorisation, Protection, Remote Working"),
        ("Related Policy", "ISMS-POL-A.7.8-9"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Semi-Annual"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "1. Complete Sheet 2 (Equipment Inventory) - categories of off-premises equipment.",
        "2. Complete Sheet 3 (Authorisation & Tracking) - removal processes.",
        "3. Complete Sheet 4 (Protection Measures) - security controls.",
        "4. Complete Sheet 5 (Remote Working) - remote work scenarios.",
        "5. Complete Sheet 6 (Permanent Off-Site) - fixed installations.",
        "6. Document incidents in Sheet 7 (Incidents).",
        "7. Document evidence in Sheet 8 (Evidence Register).",
        "8. Review Sheet 9 (Summary Dashboard) for compliance scores.",
        "9. Obtain approvals in Sheet 10 (Approval Sign-Off).",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws[f"A{row+1}"] = "STATUS LEGEND"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    legend = [
        ("Symbol", "Status", "Description"),
        (f"{CHECK}", "Compliant", "Meets all policy requirements"),
        (f"{WARNING}", "Partial", "Meets some requirements, gaps identified"),
        (f"{XMARK}", "Non-Compliant", "Does not meet requirements, action needed"),
    ]

    row += 2
    for c, h in enumerate(legend[0], start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for sym, status, desc in legend[1:]:
        ws.cell(row=row, column=1, value=sym).border = styles["border"]
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center")

        if "Compliant" in status and "Non" not in status:
            s.fill = styles["status_compliant"]["fill"]
        elif "Partial" in status:
            s.fill = styles["status_partial"]["fill"]
        elif "Non-Compliant" in status:
            s.fill = styles["status_noncompliant"]["fill"]
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 50

# =============================================================================
# EQUIPMENT INVENTORY SHEET
# =============================================================================

def create_equipment_inventory_sheet(ws, styles):
    """Create Equipment Inventory sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = (
        "Equipment Inventory for Off-Premises Use\n"
        "Policy Requirement: Off-site assets should be protected"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Category ID": 12,
        "Equipment Category": 25,
        "Equipment Type": 18,
        "Total Count": 12,
        "Off-Premises Count": 16,
        "Primary Use Case": 25,
        "MDM Managed": 18,
        "Encryption Enabled": 18,
        "Remote Wipe Capable": 18,
        "GPS Tracking": 18,
        "Last Compliance Check": 18,
        "Compliance Status": 18,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data validations
    dv_type = DataValidation(
        type="list",
        formula1='"Laptop,Mobile Phone,Tablet,Storage Media,Edge Device,Other"',
        allow_blank=False
    )
    dv_mdm = DataValidation(
        type="list",
        formula1='"Yes - All devices,Partial,No"',
        allow_blank=False
    )
    dv_encryption = DataValidation(
        type="list",
        formula1='"Yes - 100%,Partial (>80%),Low (<80%),No"',
        allow_blank=False
    )
    dv_wipe = DataValidation(
        type="list",
        formula1='"Yes - Tested,Yes - Not tested,No"',
        allow_blank=False
    )
    dv_gps = DataValidation(
        type="list",
        formula1='"Yes - Enabled,Available - Not enabled,Not available"',
        allow_blank=False
    )
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant"',
        allow_blank=False
    )

    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_mdm)
    ws.add_data_validation(dv_encryption)
    ws.add_data_validation(dv_wipe)
    ws.add_data_validation(dv_gps)
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"CAT-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=3))
        dv_mdm.add(ws.cell(row=r, column=7))
        dv_encryption.add(ws.cell(row=r, column=8))
        dv_wipe.add(ws.cell(row=r, column=9))
        dv_gps.add(ws.cell(row=r, column=10))
        dv_status.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "A4"

# =============================================================================
# AUTHORISATION & TRACKING SHEET
# =============================================================================

def create_authorisation_sheet(ws, styles):
    """Create Authorisation & Tracking sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = (
        "Authorisation & Tracking Processes\n"
        "Policy Requirement: Equipment removal should be authorised and tracked"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Category ID": 12,
        "Equipment Category": 25,
        "Authorisation Required": 22,
        "Authorisation Method": 25,
        "Tracking System": 20,
        "Chain of Custody": 20,
        "Return Verification": 22,
        "Overdue Alert": 18,
        "Current Overdue": 15,
        "Last Process Review": 18,
        "Compliance Status": 18,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_auth = DataValidation(
        type="list",
        formula1='"Yes - Manager approval,Yes - IT approval,Yes - Auto-approved,No approval required"',
        allow_blank=False
    )
    dv_custody = DataValidation(
        type="list",
        formula1='"Yes - Full tracking,Partial,No"',
        allow_blank=False
    )
    dv_return = DataValidation(
        type="list",
        formula1='"Yes - Physical check,Yes - System check,No verification"',
        allow_blank=False
    )
    dv_alert = DataValidation(
        type="list",
        formula1='"Yes - Automated,Yes - Manual review,No"',
        allow_blank=False
    )
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant"',
        allow_blank=False
    )

    ws.add_data_validation(dv_auth)
    ws.add_data_validation(dv_custody)
    ws.add_data_validation(dv_return)
    ws.add_data_validation(dv_alert)
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"CAT-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_auth.add(ws.cell(row=r, column=3))
        dv_custody.add(ws.cell(row=r, column=6))
        dv_return.add(ws.cell(row=r, column=7))
        dv_alert.add(ws.cell(row=r, column=8))
        dv_status.add(ws.cell(row=r, column=11))

    ws.freeze_panes = "A4"

# =============================================================================
# PROTECTION MEASURES SHEET
# =============================================================================

def create_protection_measures_sheet(ws, styles):
    """Create Protection Measures sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = (
        "Protection Measures for Off-Premises Equipment\n"
        "Policy Requirement: Equipment should be protected when off-premises"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Category ID": 12,
        "Equipment Category": 25,
        "Physical Security": 22,
        "Transport Guidelines": 22,
        "Public Place Rules": 22,
        "Storage When Not In Use": 25,
        "Environmental Protection": 22,
        "Privacy Screen Required": 22,
        "VPN Required": 20,
        "Screen Lock Timeout": 18,
        "Compliance Status": 18,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_physical = DataValidation(
        type="list",
        formula1='"Cable lock required,Secure bag required,No requirement"',
        allow_blank=False
    )
    dv_transport = DataValidation(
        type="list",
        formula1='"Yes - Documented,Partial,No guidelines"',
        allow_blank=False
    )
    dv_public = DataValidation(
        type="list",
        formula1='"Yes - Never unattended,Yes - Partial rules,No rules"',
        allow_blank=False
    )
    dv_storage = DataValidation(
        type="list",
        formula1='"Locked storage required,Secure location recommended,No requirement"',
        allow_blank=False
    )
    dv_env = DataValidation(
        type="list",
        formula1='"Yes - Guidelines provided,Partial,No guidelines"',
        allow_blank=False
    )
    dv_privacy = DataValidation(
        type="list",
        formula1='"Yes - Always,Yes - For sensitive data,No"',
        allow_blank=False
    )
    dv_vpn = DataValidation(
        type="list",
        formula1='"Yes - All connections,Yes - Untrusted networks,No requirement"',
        allow_blank=False
    )
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant"',
        allow_blank=False
    )

    ws.add_data_validation(dv_physical)
    ws.add_data_validation(dv_transport)
    ws.add_data_validation(dv_public)
    ws.add_data_validation(dv_storage)
    ws.add_data_validation(dv_env)
    ws.add_data_validation(dv_privacy)
    ws.add_data_validation(dv_vpn)
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"CAT-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_physical.add(ws.cell(row=r, column=3))
        dv_transport.add(ws.cell(row=r, column=4))
        dv_public.add(ws.cell(row=r, column=5))
        dv_storage.add(ws.cell(row=r, column=6))
        dv_env.add(ws.cell(row=r, column=7))
        dv_privacy.add(ws.cell(row=r, column=8))
        dv_vpn.add(ws.cell(row=r, column=9))
        dv_status.add(ws.cell(row=r, column=11))

    ws.freeze_panes = "A4"

# =============================================================================
# REMOTE WORKING SHEET
# =============================================================================

def create_remote_working_sheet(ws, styles):
    """Create Remote Working sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = (
        "Remote Working Security Assessment\n"
        "Policy Requirement: Remote work environments should be secured"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Scenario ID": 12,
        "Work Scenario": 30,
        "Data Sensitivity": 22,
        "VPN Requirement": 22,
        "Privacy Screen": 18,
        "WiFi Security": 22,
        "Physical Security": 22,
        "Visitor Access": 20,
        "Bluetooth/Wireless": 20,
        "Worker Count": 12,
        "Last Review": 15,
        "Compliance Status": 18,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_sensitivity = DataValidation(
        type="list",
        formula1='"High (PII, Financial),Medium (Internal),Low (Public)"',
        allow_blank=False
    )
    dv_vpn = DataValidation(
        type="list",
        formula1='"Required - Always,Required - Sensitive only,Recommended,Not required"',
        allow_blank=False
    )
    dv_privacy = DataValidation(
        type="list",
        formula1='"Required,Recommended,Not required"',
        allow_blank=False
    )
    dv_wifi = DataValidation(
        type="list",
        formula1='"Encrypted only,VPN for public WiFi,No requirement"',
        allow_blank=False
    )
    dv_physical = DataValidation(
        type="list",
        formula1='"Locked when away,Not visible to others,No requirement"',
        allow_blank=False
    )
    dv_visitor = DataValidation(
        type="list",
        formula1='"No access allowed,Supervised access only,N/A"',
        allow_blank=False
    )
    dv_bluetooth = DataValidation(
        type="list",
        formula1='"Disable when not needed,No requirement"',
        allow_blank=False
    )
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant"',
        allow_blank=False
    )

    ws.add_data_validation(dv_sensitivity)
    ws.add_data_validation(dv_vpn)
    ws.add_data_validation(dv_privacy)
    ws.add_data_validation(dv_wifi)
    ws.add_data_validation(dv_physical)
    ws.add_data_validation(dv_visitor)
    ws.add_data_validation(dv_bluetooth)
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"RW-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_sensitivity.add(ws.cell(row=r, column=3))
        dv_vpn.add(ws.cell(row=r, column=4))
        dv_privacy.add(ws.cell(row=r, column=5))
        dv_wifi.add(ws.cell(row=r, column=6))
        dv_physical.add(ws.cell(row=r, column=7))
        dv_visitor.add(ws.cell(row=r, column=8))
        dv_bluetooth.add(ws.cell(row=r, column=9))
        dv_status.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "A4"

# =============================================================================
# PERMANENT OFF-SITE SHEET
# =============================================================================

def create_permanent_offsite_sheet(ws, styles):
    """Create Permanent Off-Site Installations sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = (
        "Permanent Off-Site Installations\n"
        "Policy Requirement: Permanently off-site equipment requires enhanced protection"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Installation ID": 12,
        "Installation Name": 25,
        "Installation Type": 20,
        "Location Count": 12,
        "Physical Security": 22,
        "Environmental Monitoring": 22,
        "Remote Management": 20,
        "Inspection Schedule": 18,
        "Last Inspection": 15,
        "Incident Response": 20,
        "Compliance Status": 18,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"ATM/Kiosk,Sensor/IoT,Edge Computing,Signage,Antenna/Telecom,Other"',
        allow_blank=False
    )
    dv_physical = DataValidation(
        type="list",
        formula1='"Tamper detection,Locked enclosure,Public exposure"',
        allow_blank=False
    )
    dv_env = DataValidation(
        type="list",
        formula1='"Yes - Continuous,Yes - Periodic,No"',
        allow_blank=False
    )
    dv_remote = DataValidation(
        type="list",
        formula1='"Yes - Full control,Partial,No"',
        allow_blank=False
    )
    dv_schedule = DataValidation(
        type="list",
        formula1='"Monthly,Quarterly,Annual,None"',
        allow_blank=False
    )
    dv_incident = DataValidation(
        type="list",
        formula1='"Yes - Documented,Partial,No procedure"',
        allow_blank=False
    )
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant"',
        allow_blank=False
    )

    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_physical)
    ws.add_data_validation(dv_env)
    ws.add_data_validation(dv_remote)
    ws.add_data_validation(dv_schedule)
    ws.add_data_validation(dv_incident)
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"PERM-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=3))
        dv_physical.add(ws.cell(row=r, column=5))
        dv_env.add(ws.cell(row=r, column=6))
        dv_remote.add(ws.cell(row=r, column=7))
        dv_schedule.add(ws.cell(row=r, column=8))
        dv_incident.add(ws.cell(row=r, column=10))
        dv_status.add(ws.cell(row=r, column=11))

    ws.freeze_panes = "A4"

# =============================================================================
# INCIDENTS SHEET
# =============================================================================

def create_incidents_sheet(ws, styles):
    """Create Incidents tracking sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = (
        "Equipment Security Incidents\n"
        "Policy Requirement: Track and learn from equipment security incidents"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Incident ID": 12,
        "Incident Date": 12,
        "Equipment Category": 20,
        "Incident Type": 15,
        "Location": 20,
        "Data at Risk": 22,
        "Remote Wipe Executed": 22,
        "Time to Report (hrs)": 18,
        "Recovery Status": 18,
        "Root Cause": 25,
        "Corrective Action": 30,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Lost,Stolen,Damaged,Compromised,Near Miss"',
        allow_blank=False
    )
    dv_data = DataValidation(
        type="list",
        formula1='"High (PII, Financial),Medium (Internal),Low (Public),None (encrypted)"',
        allow_blank=False
    )
    dv_wipe = DataValidation(
        type="list",
        formula1='"Yes - Successful,Yes - Failed,No - Not needed,No - Not possible"',
        allow_blank=False
    )
    dv_recovery = DataValidation(
        type="list",
        formula1='"Recovered,Not recovered,Insurance claim,Replaced"',
        allow_blank=False
    )

    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_data)
    ws.add_data_validation(dv_wipe)
    ws.add_data_validation(dv_recovery)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"INC-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=4))
        dv_data.add(ws.cell(row=r, column=6))
        dv_wipe.add(ws.cell(row=r, column=7))
        dv_recovery.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "A4"

# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Document all supporting evidence for audit traceability."
    ws["A2"].font = Font(name="Calibri", italic=True)

    headers = [
        "Evidence ID", "Evidence Type", "Description", "Related Sheet/Item",
        "File Name", "File Location", "Collection Date", "Collected By",
        "Retention Period", "Notes",
    ]
    widths = [12, 18, 40, 25, 35, 50, 15, 25, 18, 40]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Screenshot,MDM Report,Policy Document,Authorisation Form,Incident Report,Test Results,Other"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_type)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EVID-{r-4:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "A5"

# =============================================================================
# SUMMARY DASHBOARD
# =============================================================================

def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with metrics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "OFF-PREMISES ASSET SECURITY SUMMARY DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Period:"
    ws["A3"].font = Font(name="Calibri", bold=True)
    ws["B3"].fill = styles["input_cell"]["fill"]
    ws["B3"].border = styles["border"]

    ws["A5"] = "Compliance Metrics"
    ws["A5"].font = Font(name="Calibri", size=12, bold=True)

    headers = ["Domain", "Items", f"{CHECK} Compliant", f"{WARNING} Partial", f"{XMARK} Non-Compliant", "% Compliant"]
    row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    domains = [
        ("Equipment Inventory", "Equipment Inventory", 12),
        ("Authorisation & Tracking", "Authorisation", 11),
        ("Protection Measures", "Protection Measures", 11),
        ("Remote Working", "Remote Working", 12),
        ("Permanent Off-Site", "Permanent Off-Site", 11),
    ]

    row += 1
    start_data_row = row
    for label, sheet, status_col in domains:
        ws.cell(row=row, column=1, value=label)
        status_col_letter = get_column_letter(status_col)
        status_range = f"'{sheet}'!{status_col_letter}4:{status_col_letter}53"

        ws.cell(row=row, column=2, value=f'=COUNTA({status_range})')
        ws.cell(row=row, column=3, value=f'=COUNTIF({status_range},"{CHECK}*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({status_range},"{WARNING}*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({status_range},"{XMARK}*")')
        ws.cell(row=row, column=6, value=f'=IF(B{row}=0,"0%",ROUND(C{row}/B{row}*100,1)&"%")')
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True)
    for col in range(2, 6):
        cell = ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{row-1})")
        cell.font = Font(name="Calibri", bold=True)

    total_pct = ws.cell(row=row, column=6, value=f'=IF(B{row}=0,"0%",ROUND(C{row}/B{row}*100,1)&"%")')
    total_pct.font = Font(name="Calibri", bold=True, color="0000FF", size=12)

    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 22

    # Incident metrics
    row += 3
    ws[f"A{row}"] = "Incident Summary"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)

    row += 1
    ws[f"A{row}"] = "Total Incidents (This Period):"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws[f"B{row}"] = "=COUNTA(Incidents!A4:A103)"

    row += 1
    ws[f"A{row}"] = "Devices Lost/Stolen:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws[f"B{row}"] = '=COUNTIF(Incidents!D4:D103,"Lost")+COUNTIF(Incidents!D4:D103,"Stolen")'

    row += 1
    ws[f"A{row}"] = "Remote Wipe Success Rate:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws[f"B{row}"] = '=IF(COUNTA(Incidents!G4:G103)=0,"N/A",ROUND(COUNTIF(Incidents!G4:G103,"Yes - Successful")/COUNTA(Incidents!G4:G103)*100,1)&"%")'

# =============================================================================
# APPROVAL SIGN-OFF
# =============================================================================

def create_approval_signoff(ws, styles):
    """Create Approval and Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Summary"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    summary_fields = [
        ("Assessment Document", "ISMS-IMP-A.7.9.S2 - Off-Premises Asset Security"),
        ("Assessment Period", ""),
        ("Overall Compliance Rate", "='Summary Dashboard'!F11"),
        ("Assessment Status", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Approval Workflow"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    approval_headers = ["Level", "Role", "Name", "Date", "Signature"]
    row += 3
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    roles = [
        ("Level 1", "Assessor"),
        ("Level 2", "IT Operations Manager"),
        ("Level 3", "CISO"),
        ("Level 4", "Compliance Officer"),
    ]
    row += 1
    for level, role in roles:
        ws.cell(row=row, column=1, value=level).font = Font(name="Calibri", bold=True)
        ws.cell(row=row, column=2, value=role).font = Font(name="Calibri", bold=True)
        for col in range(3, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 30

# =============================================================================
# MAIN WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    styles = setup_styles()

    # Create all sheets
    ws = wb.create_sheet("Instructions & Legend", 0)
    create_instructions_sheet(ws, styles)

    ws = wb.create_sheet("Equipment Inventory")
    create_equipment_inventory_sheet(ws, styles)

    ws = wb.create_sheet("Authorisation")
    create_authorisation_sheet(ws, styles)

    ws = wb.create_sheet("Protection Measures")
    create_protection_measures_sheet(ws, styles)

    ws = wb.create_sheet("Remote Working")
    create_remote_working_sheet(ws, styles)

    ws = wb.create_sheet("Permanent Off-Site")
    create_permanent_offsite_sheet(ws, styles)

    ws = wb.create_sheet("Incidents")
    create_incidents_sheet(ws, styles)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)

    ws = wb.create_sheet("Summary Dashboard")
    create_summary_dashboard(ws, styles)

    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_signoff(ws, styles)

    return wb

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main entry point for workbook generation."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS Off-Premises Asset Security Assessment (A.7.9)")
        logger.info("=" * 70)

        wb = create_workbook()
        filename = OUTPUT_FILENAME
        wb.save(filename)

        logger.info("%s SUCCESS: %s", CHECK, filename)
        logger.info("  %s 10 professional worksheets created", BULLET)
        logger.info("  %s Equipment Inventory, Authorisation, Protection Measures", BULLET)
        logger.info("  %s Remote Working, Permanent Off-Site, Incidents", BULLET)
        logger.info("  %s Evidence Register, Summary Dashboard, Approval Sign-Off", BULLET)
        logger.info("  %s 50-100 data rows per assessment sheet", BULLET)
        logger.info("  %s Data validations and dropdowns configured", BULLET)
        logger.info("  %s Compliance and incident formulas in Summary Dashboard", BULLET)
        logger.info("=" * 70)

        return 0

    except Exception as e:
        logger.error("%s ERROR: Failed to generate workbook", XMARK)
        logger.error("Error: %s", str(e))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
