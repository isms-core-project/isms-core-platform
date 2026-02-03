#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.8.S1 - Equipment Siting Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.8: Equipment Siting and Protection
Assessment Domain 1 of 3: Equipment Siting Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific equipment siting requirements, facility types,
and assessment criteria.

Key customization areas:
1. Location types and criticality tiers (match your facility classifications)
2. Access control types (adapt to your security systems)
3. Environmental thresholds (based on your equipment requirements)
4. Compliance formulas (aligned with your policy requirements)

Reference Pattern: Based on ISMS-A.7.4 Physical Infrastructure Framework

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Usage:
    python3 generate_a78_1_equipment_siting.py

Output:
    ISMS-IMP-A.7.8.S1_Equipment_Siting_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.8.S1"
WORKBOOK_NAME = "Equipment Siting Assessment"
CONTROL_ID = "A.7.8"
CONTROL_NAME = "Equipment Siting and Protection"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # Swiss format
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = '\u2705'      # Green checkmark
XMARK = '\u274C'      # Red X
WARNING = '\u26A0'    # Warning sign
BULLET = '\u2022'     # Bullet point

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.7.8.S1 - Equipment Siting Assessment\n"
        "ISO/IEC 27001:2022 - Control A.7.8: Equipment Siting and Protection"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.7.8.S1"),
        ("Assessment Area", "Equipment Siting - Location, Environment, Security, Power"),
        ("Related Policy", "ISMS-POL-A.7.8-9"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Annual"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "1. Complete Sheet 2 (Equipment Locations) - inventory all locations with equipment.",
        "2. Complete Sheet 3 (Environmental Assessment) for each location.",
        "3. Complete Sheet 4 (Physical Security) for each location.",
        "4. Complete Sheet 5 (Power Infrastructure) for each location.",
        "5. Complete Sheet 6 (Workstation Security) for office areas.",
        "6. Document evidence in Sheet 7 (Evidence Register).",
        "7. Review Sheet 8 (Summary Dashboard) for compliance scores.",
        "8. Obtain approvals in Sheet 9 (Approval Sign-Off).",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws[f"A{row+1}"] = "STATUS LEGEND"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    legend = [
        ("Symbol", "Status", "Description"),
        (f"{CHECK}", "Compliant", "Meets all policy requirements"),
        (f"{WARNING}", "Partial", "Meets some requirements, gaps identified"),
        (f"{XMARK}", "Non-Compliant", "Does not meet requirements, action needed"),
        ("N/A", "Not Applicable", "Not required for this location"),
    ]

    row += 2
    for c, h in enumerate(legend[0], start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for sym, status, desc in legend[1:]:
        ws.cell(row=row, column=1, value=sym).border = styles["border"]
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if "Compliant" in status and "Non" not in status:
            s.fill = styles["status_compliant"]["fill"]
        elif "Partial" in status:
            s.fill = styles["status_partial"]["fill"]
        elif "Non-Compliant" in status:
            s.fill = styles["status_noncompliant"]["fill"]
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 50

# =============================================================================
# EQUIPMENT LOCATIONS SHEET
# =============================================================================

def create_equipment_locations_sheet(ws, styles):
    """Create Equipment Locations assessment sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = (
        "Equipment Locations Inventory\n"
        "Policy Requirement: Equipment should be sited securely and protected"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Location ID": 12,
        "Location Name": 30,
        "Location Type": 18,
        "Building/Address": 35,
        "Equipment Types": 30,
        "Criticality Tier": 18,
        "Access Control Type": 20,
        "Environmental Monitoring": 22,
        "CCTV Coverage": 18,
        "UPS Protected": 18,
        "Last Inspection": 15,
        "Compliance Status": 18,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data validation dropdowns
    dv_location_type = DataValidation(
        type="list",
        formula1='"Data Centre,Server Room,Network Closet,Office Area,Co-location,Remote Office,Other"',
        allow_blank=False
    )
    dv_criticality = DataValidation(
        type="list",
        formula1='"Tier 1 - Critical,Tier 2 - Standard,Tier 3 - Non-Critical"',
        allow_blank=False
    )
    dv_access = DataValidation(
        type="list",
        formula1='"Biometric + Badge,Badge Only,Key Lock,Combination Lock,No Access Control"',
        allow_blank=False
    )
    dv_env_mon = DataValidation(
        type="list",
        formula1='"Yes - Real-time alerts,Yes - Periodic checks,No monitoring"',
        allow_blank=False
    )
    dv_cctv = DataValidation(
        type="list",
        formula1='"Yes - 24/7 recording,Yes - Motion triggered,No CCTV"',
        allow_blank=False
    )
    dv_ups = DataValidation(
        type="list",
        formula1='"Yes - Full coverage,Partial coverage,No UPS"',
        allow_blank=False
    )
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=False
    )

    ws.add_data_validation(dv_location_type)
    ws.add_data_validation(dv_criticality)
    ws.add_data_validation(dv_access)
    ws.add_data_validation(dv_env_mon)
    ws.add_data_validation(dv_cctv)
    ws.add_data_validation(dv_ups)
    ws.add_data_validation(dv_status)

    # Data rows
    for r in range(4, 104):
        # Pre-fill Location ID
        ws.cell(row=r, column=1, value=f"LOC-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Apply dropdowns
        dv_location_type.add(ws.cell(row=r, column=3))
        dv_criticality.add(ws.cell(row=r, column=6))
        dv_access.add(ws.cell(row=r, column=7))
        dv_env_mon.add(ws.cell(row=r, column=8))
        dv_cctv.add(ws.cell(row=r, column=9))
        dv_ups.add(ws.cell(row=r, column=10))
        dv_status.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "A4"

# =============================================================================
# ENVIRONMENTAL ASSESSMENT SHEET
# =============================================================================

def create_environmental_sheet(ws, styles):
    """Create Environmental Assessment sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = (
        "Environmental Assessment\n"
        "Policy Requirement: Equipment should be protected from environmental threats"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Location ID": 12,
        "Location Name": 30,
        "Temperature Range": 18,
        "Humidity Range": 18,
        "Temperature Monitoring": 22,
        "Flood Risk": 18,
        "Fire Suppression": 22,
        "Dust/Contamination": 18,
        "Vibration Exposure": 18,
        "Food/Drink Prohibited": 20,
        "Incidents (12 months)": 18,
        "Compliance Status": 18,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data validation
    dv_temp_mon = DataValidation(
        type="list",
        formula1='"Yes - Real-time alerts,Yes - Daily checks,No monitoring"',
        allow_blank=False
    )
    dv_flood = DataValidation(
        type="list",
        formula1='"High (below grade),Medium (ground floor),Low (upper floor),Protected (raised floor)"',
        allow_blank=False
    )
    dv_fire = DataValidation(
        type="list",
        formula1='"Gas suppression,Pre-action sprinkler,Wet sprinkler,Extinguishers only,None"',
        allow_blank=False
    )
    dv_dust = DataValidation(
        type="list",
        formula1='"Clean room,Filtered HVAC,Standard office,Industrial/Dusty"',
        allow_blank=False
    )
    dv_vibration = DataValidation(
        type="list",
        formula1='"Low (isolated),Medium (standard building),High (near machinery)"',
        allow_blank=False
    )
    dv_food = DataValidation(
        type="list",
        formula1='"Yes - Enforced,Yes - Policy only,No restriction"',
        allow_blank=False
    )
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=False
    )

    ws.add_data_validation(dv_temp_mon)
    ws.add_data_validation(dv_flood)
    ws.add_data_validation(dv_fire)
    ws.add_data_validation(dv_dust)
    ws.add_data_validation(dv_vibration)
    ws.add_data_validation(dv_food)
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"LOC-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_temp_mon.add(ws.cell(row=r, column=5))
        dv_flood.add(ws.cell(row=r, column=6))
        dv_fire.add(ws.cell(row=r, column=7))
        dv_dust.add(ws.cell(row=r, column=8))
        dv_vibration.add(ws.cell(row=r, column=9))
        dv_food.add(ws.cell(row=r, column=10))
        dv_status.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "A4"

# =============================================================================
# PHYSICAL SECURITY SHEET
# =============================================================================

def create_physical_security_sheet(ws, styles):
    """Create Physical Security assessment sheet."""
    ws.merge_cells("A1:N1")
    ws["A1"] = (
        "Physical Security Assessment\n"
        "Policy Requirement: Equipment should be protected by appropriate physical security measures"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Location ID": 12,
        "Location Name": 25,
        "Access Control System": 25,
        "Access Levels Defined": 20,
        "Log Retention (Days)": 18,
        "CCTV System": 22,
        "CCTV Retention (Days)": 18,
        "Equipment Locking": 22,
        "Cable Protection": 18,
        "Asset Labels": 16,
        "Segregation": 22,
        "Last Security Review": 15,
        "Compliance Status": 18,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_locking = DataValidation(
        type="list",
        formula1='"Rack locks + Cable locks,Rack locks only,Cable locks only,None"',
        allow_blank=False
    )
    dv_cable = DataValidation(
        type="list",
        formula1='"Enclosed conduit,Cable trays,Exposed cabling"',
        allow_blank=False
    )
    dv_labels = DataValidation(
        type="list",
        formula1='"Yes - All labelled,Partial,No labels"',
        allow_blank=False
    )
    dv_seg = DataValidation(
        type="list",
        formula1='"Yes - Physical separation,Partial - Logical only,No - Shared space"',
        allow_blank=False
    )
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=False
    )

    ws.add_data_validation(dv_locking)
    ws.add_data_validation(dv_cable)
    ws.add_data_validation(dv_labels)
    ws.add_data_validation(dv_seg)
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"LOC-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_locking.add(ws.cell(row=r, column=8))
        dv_cable.add(ws.cell(row=r, column=9))
        dv_labels.add(ws.cell(row=r, column=10))
        dv_seg.add(ws.cell(row=r, column=11))
        dv_status.add(ws.cell(row=r, column=13))

    ws.freeze_panes = "A4"

# =============================================================================
# POWER INFRASTRUCTURE SHEET
# =============================================================================

def create_power_infrastructure_sheet(ws, styles):
    """Create Power Infrastructure assessment sheet."""
    ws.merge_cells("A1:N1")
    ws["A1"] = (
        "Power Infrastructure Assessment\n"
        "Policy Requirement: Equipment should be protected from power failures"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Location ID": 12,
        "Location Name": 25,
        "UPS Coverage": 22,
        "UPS Runtime (min)": 18,
        "Generator Backup": 22,
        "Surge Protection": 18,
        "Lightning Protection": 20,
        "EPO Switch": 18,
        "Power Redundancy": 22,
        "Last UPS Test": 15,
        "Last Generator Test": 18,
        "Power Incidents (12m)": 18,
        "Compliance Status": 18,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_ups = DataValidation(
        type="list",
        formula1='"100% (all equipment),Partial (critical only),None"',
        allow_blank=False
    )
    dv_gen = DataValidation(
        type="list",
        formula1='"Yes - Auto-transfer,Yes - Manual transfer,No generator"',
        allow_blank=False
    )
    dv_surge = DataValidation(
        type="list",
        formula1='"Yes - All circuits,Partial,None"',
        allow_blank=False
    )
    dv_lightning = DataValidation(
        type="list",
        formula1='"Yes - Building grounded,Unknown,No"',
        allow_blank=False
    )
    dv_epo = DataValidation(
        type="list",
        formula1='"Yes - Near room,Yes - Remote only,No EPO"',
        allow_blank=False
    )
    dv_redundancy = DataValidation(
        type="list",
        formula1='"Dual feed (A+B),Single feed + UPS,Single feed only"',
        allow_blank=False
    )
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=False
    )

    ws.add_data_validation(dv_ups)
    ws.add_data_validation(dv_gen)
    ws.add_data_validation(dv_surge)
    ws.add_data_validation(dv_lightning)
    ws.add_data_validation(dv_epo)
    ws.add_data_validation(dv_redundancy)
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"LOC-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_ups.add(ws.cell(row=r, column=3))
        dv_gen.add(ws.cell(row=r, column=5))
        dv_surge.add(ws.cell(row=r, column=6))
        dv_lightning.add(ws.cell(row=r, column=7))
        dv_epo.add(ws.cell(row=r, column=8))
        dv_redundancy.add(ws.cell(row=r, column=9))
        dv_status.add(ws.cell(row=r, column=13))

    ws.freeze_panes = "A4"

# =============================================================================
# WORKSTATION SECURITY SHEET
# =============================================================================

def create_workstation_security_sheet(ws, styles):
    """Create Workstation Security assessment sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = (
        "Workstation Security Assessment\n"
        "Policy Requirement: Screens should be positioned to prevent unauthorised viewing"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Area ID": 12,
        "Area Name": 30,
        "Workstation Count": 15,
        "Data Sensitivity": 22,
        "Screen Positioning": 25,
        "Privacy Screens": 20,
        "Automatic Lock": 20,
        "Clear Desk Policy": 20,
        "Visitor Access": 18,
        "Shoulder Surfing Risk": 20,
        "Last Review": 15,
        "Compliance Status": 18,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_sensitivity = DataValidation(
        type="list",
        formula1='"High (PII, Financial),Medium (Internal),Low (Public)"',
        allow_blank=False
    )
    dv_positioning = DataValidation(
        type="list",
        formula1='"Optimal (away from windows/traffic),Acceptable,At risk (visible to public)"',
        allow_blank=False
    )
    dv_privacy = DataValidation(
        type="list",
        formula1='"Yes - All workstations,Partial,None"',
        allow_blank=False
    )
    dv_lock = DataValidation(
        type="list",
        formula1='"Yes - <5 min timeout,Yes - >5 min timeout,Not configured"',
        allow_blank=False
    )
    dv_desk = DataValidation(
        type="list",
        formula1='"Yes - Enforced,Yes - Policy only,No policy"',
        allow_blank=False
    )
    dv_visitor = DataValidation(
        type="list",
        formula1='"Escorted only,Badge required,Open access"',
        allow_blank=False
    )
    dv_shoulder = DataValidation(
        type="list",
        formula1='"Low,Medium,High"',
        allow_blank=False
    )
    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=False
    )

    ws.add_data_validation(dv_sensitivity)
    ws.add_data_validation(dv_positioning)
    ws.add_data_validation(dv_privacy)
    ws.add_data_validation(dv_lock)
    ws.add_data_validation(dv_desk)
    ws.add_data_validation(dv_visitor)
    ws.add_data_validation(dv_shoulder)
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"WS-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_sensitivity.add(ws.cell(row=r, column=4))
        dv_positioning.add(ws.cell(row=r, column=5))
        dv_privacy.add(ws.cell(row=r, column=6))
        dv_lock.add(ws.cell(row=r, column=7))
        dv_desk.add(ws.cell(row=r, column=8))
        dv_visitor.add(ws.cell(row=r, column=9))
        dv_shoulder.add(ws.cell(row=r, column=10))
        dv_status.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "A4"

# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Document all supporting evidence for audit traceability."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID",
        "Evidence Type",
        "Description",
        "Related Sheet/Item",
        "File Name",
        "File Location",
        "Collection Date",
        "Collected By",
        "Retention Period",
        "Notes",
    ]
    widths = [12, 18, 40, 25, 35, 50, 15, 25, 18, 40]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Screenshot,Configuration Export,Log Sample,Report,Document,Photo,Floor Plan"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_type)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EVID-{r-4:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "A5"

# =============================================================================
# SUMMARY DASHBOARD
# =============================================================================

def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with metrics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "EQUIPMENT SITING SUMMARY DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Period:"
    ws["A3"].font = Font(name="Calibri", bold=True)
    ws["B3"].fill = styles["input_cell"]["fill"]
    ws["B3"].border = styles["border"]

    ws["A5"] = "Compliance Metrics"
    ws["A5"].font = Font(name="Calibri", size=12, bold=True)

    headers = ["Domain", "Items", f"{CHECK} Compliant", f"{WARNING} Partial", f"{XMARK} Non-Compliant", "N/A", "% Compliant"]
    row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    domains = [
        ("Equipment Locations", "Equipment Locations", 12),
        ("Environmental", "Environmental", 12),
        ("Physical Security", "Physical Security", 13),
        ("Power Infrastructure", "Power Infrastructure", 13),
        ("Workstation Security", "Workstation Security", 12),
    ]

    row += 1
    start_data_row = row
    for label, sheet, status_col in domains:
        ws.cell(row=row, column=1, value=label)
        status_col_letter = get_column_letter(status_col)
        status_range = f"'{sheet}'!{status_col_letter}4:{status_col_letter}103"

        ws.cell(row=row, column=2, value=f'=COUNTA({status_range})')
        ws.cell(row=row, column=3, value=f'=COUNTIF({status_range},"{CHECK}*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({status_range},"{WARNING}*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({status_range},"{XMARK}*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({status_range},"N/A")')
        ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{row-1})")
        cell.font = Font(name="Calibri", bold=True)

    total_pct = ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
    total_pct.font = Font(name="Calibri", bold=True, color="0000FF", size=12)

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 18

# =============================================================================
# APPROVAL SIGN-OFF
# =============================================================================

def create_approval_signoff(ws, styles):
    """Create Approval and Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Summary"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    summary_fields = [
        ("Assessment Document", "ISMS-IMP-A.7.8.S1 - Equipment Siting Assessment"),
        ("Assessment Period", ""),
        ("Overall Compliance Rate", "='Summary Dashboard'!G11"),
        ("Assessment Status", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Approval Workflow"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    approval_headers = ["Level", "Role", "Name", "Date", "Signature"]
    row += 3
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    roles = [
        ("Level 1", "Assessor"),
        ("Level 2", "Facilities Manager"),
        ("Level 3", "CISO"),
        ("Level 4", "Compliance Officer"),
    ]
    row += 1
    for level, role in roles:
        ws.cell(row=row, column=1, value=level).font = Font(name="Calibri", bold=True)
        ws.cell(row=row, column=2, value=role).font = Font(name="Calibri", bold=True)
        for col in range(3, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 30

# =============================================================================
# MAIN WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    styles = setup_styles()

    # Create all sheets
    ws = wb.create_sheet("Instructions & Legend", 0)
    create_instructions_sheet(ws, styles)

    ws = wb.create_sheet("Equipment Locations")
    create_equipment_locations_sheet(ws, styles)

    ws = wb.create_sheet("Environmental")
    create_environmental_sheet(ws, styles)

    ws = wb.create_sheet("Physical Security")
    create_physical_security_sheet(ws, styles)

    ws = wb.create_sheet("Power Infrastructure")
    create_power_infrastructure_sheet(ws, styles)

    ws = wb.create_sheet("Workstation Security")
    create_workstation_security_sheet(ws, styles)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)

    ws = wb.create_sheet("Summary Dashboard")
    create_summary_dashboard(ws, styles)

    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_signoff(ws, styles)

    return wb

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main entry point for workbook generation."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS Equipment Siting Assessment (A.7.8)")
        logger.info("=" * 70)

        wb = create_workbook()
        filename = OUTPUT_FILENAME
        wb.save(filename)

        logger.info("%s SUCCESS: %s", CHECK, filename)
        logger.info("  %s 9 professional worksheets created", BULLET)
        logger.info("  %s Equipment Locations, Environmental, Physical Security", BULLET)
        logger.info("  %s Power Infrastructure, Workstation Security", BULLET)
        logger.info("  %s Evidence Register, Summary Dashboard, Approval Sign-Off", BULLET)
        logger.info("  %s 100 data rows per assessment sheet", BULLET)
        logger.info("  %s Data validations and dropdowns configured", BULLET)
        logger.info("  %s Compliance formulas in Summary Dashboard", BULLET)
        logger.info("=" * 70)

        return 0

    except Exception as e:
        logger.error("%s ERROR: Failed to generate workbook", XMARK)
        logger.error("Error: %s", str(e))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
