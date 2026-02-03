#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.12-13.S1 - Cabling Security Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.12: Cabling Security
Assessment Domain 1 of 4: Cabling Security Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates the Cabling Security Assessment workbook for documenting
cable pathway protection, physical security measures, access controls, and
documentation compliance.

Reference Pattern: Based on ISMS-A.7.4-5-11 Physical Infrastructure Framework

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    pip3 install openpyxl

Usage:
    python3 generate_a712_1_cabling_security.py

Output:
    ISMS-IMP-A.7.12-13.S1_Cabling_Security_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.12-13.S1"
WORKBOOK_NAME = "Cabling Security Assessment"
CONTROL_ID = "A.7.12"
CONTROL_NAME = "Cabling Security"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # Swiss format
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = '\u2705'      # Green checkmark
XMARK = '\u274C'      # Red X
WARNING = '\u26A0'    # Warning sign
BULLET = '\u2022'     # Bullet point

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Cabling Security - Pathways, Protection, Access Controls, Documentation"),
        ("Related Policy", "ISMS-POL-A.7.12-13, Section 2.1"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Annual"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "1. Complete Sheet 2 (Cable Pathways) - Document all cable routes and protection status.",
        "2. Complete Sheet 3 (Physical Protection) - Assess physical and environmental protection.",
        "3. Complete Sheet 4 (Access Controls) - Document access controls for cabling infrastructure.",
        "4. Complete Sheet 5 (Documentation) - Audit cable documentation compliance.",
        "5. Review Sheet 6 (Summary Dashboard) - Automated compliance scoring.",
        "6. Complete Sheet 7 (Evidence Register) - Document supporting evidence.",
        "7. Obtain approvals in Sheet 8 (Approval Sign-Off).",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws[f"A{row+1}"] = "STATUS LEGEND"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    legend = [
        ("Symbol", "Status", "Description"),
        (CHECK, "Compliant", "Meets all policy requirements"),
        (WARNING, "Partial", "Meets some requirements, gaps identified"),
        (XMARK, "Non-Compliant", "Does not meet requirements"),
        ("N/A", "Not Applicable", "Not required for this area"),
    ]

    row += 2
    for c, h in enumerate(legend[0], start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for sym, status, desc in legend[1:]:
        ws.cell(row=row, column=1, value=sym).border = styles["border"]
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if "Compliant" in status and "Non" not in status:
            s.fill = styles["status_compliant"]["fill"]
        elif "Partial" in status:
            s.fill = styles["status_partial"]["fill"]
        elif "Non-Compliant" in status:
            s.fill = styles["status_noncompliant"]["fill"]
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 40

# =============================================================================
# CABLE PATHWAYS SHEET
# =============================================================================

def create_cable_pathways_sheet(ws, styles):
    """Create Cable Pathways assessment sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = "Cable Pathways Assessment\nPolicy Requirement: Cables should be protected from interception, interference or damage"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = [
        ("Pathway ID", 12),
        ("Facility/Building", 25),
        ("Pathway Type", 18),
        ("Start Location", 25),
        ("End Location", 25),
        ("Cable Types", 20),
        ("Protection Type", 20),
        ("Length (m)", 12),
        ("Segregation Compliant", 18),
        ("Documentation Current", 18),
        ("Last Inspection", 15),
        ("Compliance Status", 18),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data validation lists
    dv_pathway_type = DataValidation(
        type="list",
        formula1='"Conduit,Cable Tray,Raised Floor,Underground,Ceiling Void,Wall Chase"',
        allow_blank=True
    )
    ws.add_data_validation(dv_pathway_type)

    dv_protection = DataValidation(
        type="list",
        formula1='"Enclosed - Metal,Enclosed - Plastic,Open Tray,Armoured,None"',
        allow_blank=True
    )
    ws.add_data_validation(dv_protection)

    dv_yes_no_na = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_yes_no_na)

    dv_yes_partial_no = DataValidation(
        type="list",
        formula1='"Yes,Partial,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_yes_partial_no)

    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    # Data rows
    for r in range(4, 104):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Apply validations
        dv_pathway_type.add(ws.cell(row=r, column=3))
        dv_protection.add(ws.cell(row=r, column=7))
        dv_yes_no_na.add(ws.cell(row=r, column=9))
        dv_yes_partial_no.add(ws.cell(row=r, column=10))
        dv_status.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "A4"

# =============================================================================
# PHYSICAL PROTECTION SHEET
# =============================================================================

def create_physical_protection_sheet(ws, styles):
    """Create Physical Protection assessment sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "Physical Protection Assessment\nPolicy Requirement: Protection from electromagnetic interference, physical damage, water, and heat"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = [
        ("Area ID", 12),
        ("Facility/Building", 25),
        ("Area Name", 25),
        ("EMI Protection", 15),
        ("Physical Damage Protection", 22),
        ("Water Protection", 15),
        ("Heat Protection", 15),
        ("Cable Route Risk Level", 18),
        ("Fibre for High Security", 18),
        ("Compliance Status", 18),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data validations
    dv_emi = DataValidation(
        type="list",
        formula1='"Yes - Shielded,Partial,No,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_emi)

    dv_physical = DataValidation(
        type="list",
        formula1='"Yes - Armoured,Yes - Enclosed,Partial,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_physical)

    dv_yes_partial_no_na = DataValidation(
        type="list",
        formula1='"Yes,Partial,No,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_yes_partial_no_na)

    dv_risk = DataValidation(
        type="list",
        formula1='"Low,Medium,High"',
        allow_blank=True
    )
    ws.add_data_validation(dv_risk)

    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_emi.add(ws.cell(row=r, column=4))
        dv_physical.add(ws.cell(row=r, column=5))
        dv_yes_partial_no_na.add(ws.cell(row=r, column=6))
        dv_yes_partial_no_na.add(ws.cell(row=r, column=7))
        dv_risk.add(ws.cell(row=r, column=8))
        dv_yes_partial_no_na.add(ws.cell(row=r, column=9))
        dv_status.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "A4"

# =============================================================================
# ACCESS CONTROLS SHEET
# =============================================================================

def create_access_controls_sheet(ws, styles):
    """Create Access Controls assessment sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "Access Controls Assessment\nPolicy Requirement: Cabling infrastructure should be in secured areas with access controls"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = [
        ("Location ID", 12),
        ("Facility/Building", 25),
        ("Infrastructure Type", 20),
        ("Location Description", 30),
        ("Lock Type", 22),
        ("Access Restricted", 25),
        ("Access Logged", 18),
        ("Occupied Monitoring", 18),
        ("Last Access Review", 18),
        ("Compliance Status", 18),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_infra_type = DataValidation(
        type="list",
        formula1='"Wiring Closet,Patch Panel,Distribution Frame,Manhole,Duct Access"',
        allow_blank=True
    )
    ws.add_data_validation(dv_infra_type)

    dv_lock_type = DataValidation(
        type="list",
        formula1='"Electronic Access Card,Keyed Lock,Combination Lock,No Lock,Cage/Enclosure"',
        allow_blank=True
    )
    ws.add_data_validation(dv_lock_type)

    dv_access_restricted = DataValidation(
        type="list",
        formula1='"Yes - IT Only,Yes - Facilities Only,Yes - Authorised Personnel,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_access_restricted)

    dv_access_logged = DataValidation(
        type="list",
        formula1='"Yes - Electronic,Yes - Manual,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_access_logged)

    dv_monitoring = DataValidation(
        type="list",
        formula1='"Yes - CCTV,Yes - Guards,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_monitoring)

    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_infra_type.add(ws.cell(row=r, column=3))
        dv_lock_type.add(ws.cell(row=r, column=5))
        dv_access_restricted.add(ws.cell(row=r, column=6))
        dv_access_logged.add(ws.cell(row=r, column=7))
        dv_monitoring.add(ws.cell(row=r, column=8))
        dv_status.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "A4"

# =============================================================================
# DOCUMENTATION SHEET
# =============================================================================

def create_documentation_sheet(ws, styles):
    """Create Documentation compliance sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Documentation Compliance Assessment\nPolicy Requirement: Cable infrastructure should be documented with cable schedules and diagrams"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = [
        ("Document Type", 22),
        ("Document Name", 35),
        ("Document Location", 40),
        ("Owner", 25),
        ("Last Updated", 15),
        ("Review Cycle", 15),
        ("Current", 12),
        ("Accessible", 18),
        ("Compliance Status", 18),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_doc_type = DataValidation(
        type="list",
        formula1='"Cable Schedule,Network Diagram,Labelling Standard,Change Log,Audit Report,Procedure Document"',
        allow_blank=True
    )
    ws.add_data_validation(dv_doc_type)

    dv_review_cycle = DataValidation(
        type="list",
        formula1='"Annual,Quarterly,On Change"',
        allow_blank=True
    )
    ws.add_data_validation(dv_review_cycle)

    dv_yes_no = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_yes_no)

    dv_accessible = DataValidation(
        type="list",
        formula1='"Yes - Online,Yes - Restricted,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_accessible)

    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_doc_type.add(ws.cell(row=r, column=1))
        dv_review_cycle.add(ws.cell(row=r, column=6))
        dv_yes_no.add(ws.cell(row=r, column=7))
        dv_accessible.add(ws.cell(row=r, column=8))
        dv_status.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "A4"

# =============================================================================
# SUMMARY DASHBOARD
# =============================================================================

def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with metrics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "CABLING SECURITY SUMMARY DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Period:"
    ws["A3"].font = Font(name="Calibri", bold=True)
    ws["B3"].fill = styles["input_cell"]["fill"]
    ws["B3"].border = styles["border"]

    ws["A5"] = "Compliance Metrics"
    ws["A5"].font = Font(name="Calibri", size=12, bold=True)

    headers = ["Domain", "Total Items", f"{CHECK} Compliant", f"{WARNING} Partial", f"{XMARK} Non-Compliant", "N/A", "% Compliant"]
    row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    domains = [
        ("Cable Pathways", "Cable Pathways", 12),
        ("Physical Protection", "Physical Protection", 10),
        ("Access Controls", "Access Controls", 10),
        ("Documentation", "Documentation", 9),
    ]

    row += 1
    start_data_row = row
    for label, sheet, status_col in domains:
        ws.cell(row=row, column=1, value=label)

        status_col_letter = get_column_letter(status_col)
        status_range = f"'{sheet}'!{status_col_letter}4:{status_col_letter}103"

        ws.cell(row=row, column=2, value=f'=COUNTA({status_range})')
        ws.cell(row=row, column=3, value=f'=COUNTIF({status_range},"{CHECK}*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({status_range},"{WARNING}*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({status_range},"{XMARK}*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({status_range},"N/A")')
        ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"N/A",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{row-1})")
        cell.font = Font(name="Calibri", bold=True)

    total_pct = ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"N/A",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
    total_pct.font = Font(name="Calibri", bold=True, color="0000FF", size=12)

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 18

# =============================================================================
# EVIDENCE REGISTER
# =============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Document all supporting evidence for audit traceability."
    ws["A2"].font = Font(name="Calibri", italic=True)

    headers = [
        ("Evidence ID", 12),
        ("Evidence Type", 18),
        ("Description", 40),
        ("Related Sheet/Item", 25),
        ("File Name", 30),
        ("File Location", 45),
        ("Collection Date", 15),
        ("Collected By", 20),
        ("Retention Period", 15),
        ("Notes", 40),
    ]

    row = 4
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Photo,Screenshot,Document,Export,Diagram,Report,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EVID-{r-4:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "A5"

# =============================================================================
# APPROVAL SIGN-OFF
# =============================================================================

def create_approval_signoff(ws, styles):
    """Create Approval and Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Summary"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Overall Compliance Rate", "='Summary Dashboard'!G11"),
        ("Assessment Status", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "" or value.startswith("="):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Approval Workflow"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    approval_headers = ["Role", "Name", "Signature", "Date", "Comments"]
    row += 3
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    roles = ["Assessor", "Facilities Manager", "CISO", "Compliance Officer"]
    row += 1
    for role in roles:
        ws.cell(row=row, column=1, value=role).font = Font(name="Calibri", bold=True)
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 40

# =============================================================================
# MAIN WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    styles = setup_styles()

    # Create all sheets
    ws = wb.create_sheet("Instructions & Legend", 0)
    create_instructions_sheet(ws, styles)

    ws = wb.create_sheet("Cable Pathways")
    create_cable_pathways_sheet(ws, styles)

    ws = wb.create_sheet("Physical Protection")
    create_physical_protection_sheet(ws, styles)

    ws = wb.create_sheet("Access Controls")
    create_access_controls_sheet(ws, styles)

    ws = wb.create_sheet("Documentation")
    create_documentation_sheet(ws, styles)

    ws = wb.create_sheet("Summary Dashboard")
    create_summary_dashboard(ws, styles)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)

    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_signoff(ws, styles)

    return wb

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main entry point for workbook generation."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS Cabling Security Assessment (A.7.12)")
        logger.info("=" * 70)

        wb = create_workbook()
        wb.save(OUTPUT_FILENAME)

        logger.info("%s SUCCESS: %s", CHECK, OUTPUT_FILENAME)
        logger.info("  %s 8 professional worksheets created", BULLET)
        logger.info("  %s Navy headers, yellow inputs styling applied", BULLET)
        logger.info("  %s 100 data rows per assessment sheet", BULLET)
        logger.info("  %s Automated compliance dashboard with formulas", BULLET)
        logger.info("  %s Data validations and freeze panes configured", BULLET)
        logger.info("  %s Evidence register with audit traceability", BULLET)
        logger.info("  %s 4-level approval workflow", BULLET)
        logger.info("=" * 70)

        return 0

    except Exception as e:
        logger.error("%s ERROR: Failed to generate workbook", XMARK)
        logger.error("Error: %s", str(e))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
