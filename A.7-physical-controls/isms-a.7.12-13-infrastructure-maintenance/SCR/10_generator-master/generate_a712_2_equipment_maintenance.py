#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.12-13.S2 - Equipment Maintenance Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.13: Equipment Maintenance
Assessment Domain 2 of 4: Equipment Maintenance Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates the Equipment Maintenance Assessment workbook for documenting
maintenance programmes, personnel verification, security controls, and remote
maintenance access.

Reference Pattern: Based on ISMS-A.7.4-5-11 Physical Infrastructure Framework

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    pip3 install openpyxl

Usage:
    python3 generate_a712_2_equipment_maintenance.py

Output:
    ISMS-IMP-A.7.12-13.S2_Equipment_Maintenance_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.12-13.S2"
WORKBOOK_NAME = "Equipment Maintenance Assessment"
CONTROL_ID = "A.7.13"
CONTROL_NAME = "Equipment Maintenance"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = '\u2705'
XMARK = '\u274C'
WARNING = '\u26A0'
BULLET = '\u2022'

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Equipment Maintenance - Programme, Personnel, Security, Remote Access"),
        ("Related Policy", "ISMS-POL-A.7.12-13, Section 2.2"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "1. Complete Sheet 2 (Equipment Inventory) - Document all equipment in maintenance programme.",
        "2. Complete Sheet 3 (Maintenance Programme) - Assess programme compliance.",
        "3. Complete Sheet 4 (Personnel & Vendors) - Document authorised personnel and vendors.",
        "4. Complete Sheet 5 (Security During Maintenance) - Assess security controls.",
        "5. Complete Sheet 6 (Remote Maintenance) - Audit remote access controls.",
        "6. Review Sheet 7 (Summary Dashboard) - Automated compliance scoring.",
        "7. Complete Sheet 8 (Evidence Register) - Document supporting evidence.",
        "8. Obtain approvals in Sheet 9 (Approval Sign-Off).",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws[f"A{row+1}"] = "STATUS LEGEND"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    legend = [
        ("Symbol", "Status", "Description"),
        (CHECK, "Compliant", "Meets all policy requirements"),
        (WARNING, "Partial", "Meets some requirements, gaps identified"),
        (XMARK, "Non-Compliant", "Does not meet requirements"),
        ("N/A", "Not Applicable", "Not required"),
    ]

    row += 2
    for c, h in enumerate(legend[0], start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for sym, status, desc in legend[1:]:
        ws.cell(row=row, column=1, value=sym).border = styles["border"]
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if "Compliant" in status and "Non" not in status:
            s.fill = styles["status_compliant"]["fill"]
        elif "Partial" in status:
            s.fill = styles["status_partial"]["fill"]
        elif "Non-Compliant" in status:
            s.fill = styles["status_noncompliant"]["fill"]
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 40

# =============================================================================
# EQUIPMENT INVENTORY SHEET
# =============================================================================

def create_equipment_inventory_sheet(ws, styles):
    """Create Equipment Inventory sheet."""
    ws.merge_cells("A1:N1")
    ws["A1"] = "Equipment Inventory\nPolicy Requirement: All equipment should be included in maintenance programme"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = [
        ("Equipment ID", 15),
        ("Equipment Type", 18),
        ("Equipment Description", 35),
        ("Location", 25),
        ("Criticality", 18),
        ("Manufacturer", 20),
        ("Manufacturer Requirement", 35),
        ("Maintenance Frequency", 18),
        ("Frequency Compliant", 15),
        ("In Maintenance Programme", 18),
        ("Last Maintenance", 15),
        ("Next Scheduled", 15),
        ("Compliance Status", 18),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_equip_type = DataValidation(
        type="list",
        formula1='"Server,Network Device,Storage,UPS,HVAC,Generator,Security System,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_equip_type)

    dv_criticality = DataValidation(
        type="list",
        formula1='"Tier 1 - Critical,Tier 2 - Standard"',
        allow_blank=True
    )
    ws.add_data_validation(dv_criticality)

    dv_frequency = DataValidation(
        type="list",
        formula1='"Monthly,Quarterly,Semi-annually,Annually"',
        allow_blank=True
    )
    ws.add_data_validation(dv_frequency)

    dv_yes_no = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_yes_no)

    dv_in_programme = DataValidation(
        type="list",
        formula1='"Yes,No,Partial"',
        allow_blank=True
    )
    ws.add_data_validation(dv_in_programme)

    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_equip_type.add(ws.cell(row=r, column=2))
        dv_criticality.add(ws.cell(row=r, column=5))
        dv_frequency.add(ws.cell(row=r, column=8))
        dv_yes_no.add(ws.cell(row=r, column=9))
        dv_in_programme.add(ws.cell(row=r, column=10))
        dv_status.add(ws.cell(row=r, column=13))

    ws.freeze_panes = "A4"

# =============================================================================
# MAINTENANCE PROGRAMME SHEET
# =============================================================================

def create_maintenance_programme_sheet(ws, styles):
    """Create Maintenance Programme assessment sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "Maintenance Programme Assessment\nPolicy Requirement: Establish and implement maintenance programme following manufacturer recommendations"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Programme Elements Assessment"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    questions = [
        ("All equipment types included in maintenance programme?", "Yes,Partial,No"),
        ("Maintenance schedules follow manufacturer recommendations?", "Yes,Partial,No"),
        ("Critical equipment prioritised for preventive maintenance?", "Yes,No"),
        ("Maintenance records maintained for all activities?", "Yes,Partial,No"),
        ("Maintenance record retention meets policy (3+ years)?", "Yes,No"),
        ("Records accessible for audit?", "Yes,Partial,No"),
        ("Preventive maintenance occurring on schedule?", "Yes,Partial,No"),
        ("Overdue maintenance tracked and escalated?", "Yes,Partial,No"),
    ]

    row = 5
    col_headers = ["Assessment Question", "Status", "Evidence Reference", "Notes"]
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for question, options in questions:
        ws.cell(row=row, column=1, value=question).border = styles["border"]

        status_cell = ws.cell(row=row, column=2)
        status_cell.fill = styles["input_cell"]["fill"]
        status_cell.border = styles["border"]

        dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(status_cell)

        for col in [3, 4]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        row += 1

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 40

# =============================================================================
# PERSONNEL & VENDORS SHEET
# =============================================================================

def create_personnel_vendors_sheet(ws, styles):
    """Create Personnel & Vendors sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "Personnel & Vendors\nPolicy Requirement: Only authorised personnel should perform maintenance"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = [
        ("Personnel/Vendor ID", 18),
        ("Name", 30),
        ("Type", 18),
        ("Equipment Types Authorised", 30),
        ("Verification Required", 20),
        ("Supervision Required", 22),
        ("Contract/Agreement", 20),
        ("NDA in Place", 12),
        ("Background Check", 15),
        ("Last Verified", 15),
        ("Compliance Status", 18),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Internal Staff,Third-Party Vendor,Contractor"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)

    dv_verification = DataValidation(
        type="list",
        formula1='"Yes - Badge/ID,Yes - Escort,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_verification)

    dv_supervision = DataValidation(
        type="list",
        formula1='"Yes - Always,Yes - Sensitive Equipment,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_supervision)

    dv_yes_no_na = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_yes_no_na)

    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=3))
        dv_verification.add(ws.cell(row=r, column=5))
        dv_supervision.add(ws.cell(row=r, column=6))
        dv_yes_no_na.add(ws.cell(row=r, column=8))
        dv_yes_no_na.add(ws.cell(row=r, column=9))
        dv_status.add(ws.cell(row=r, column=11))

    ws.freeze_panes = "A4"

# =============================================================================
# SECURITY DURING MAINTENANCE SHEET
# =============================================================================

def create_security_maintenance_sheet(ws, styles):
    """Create Security During Maintenance sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "Security During Maintenance\nPolicy Requirement: Sensitive data should be protected during maintenance activities"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Data Protection Assessment"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    sections = [
        ("Data Protection", [
            ("Sensitive data protected during maintenance?", "Yes,Partial,No"),
            ("Equipment containing data kept on-premises where possible?", "Yes,Partial,No"),
            ("Data erased before off-site maintenance?", "Yes,Partial,No,N/A"),
            ("Maintenance personnel data access restricted?", "Yes,No"),
        ]),
        ("Access Controls", [
            ("Maintenance access time-limited?", "Yes,No"),
            ("Access logged (who, what, when, why)?", "Yes,Partial,No"),
            ("Tools and equipment accounted for after maintenance?", "Yes,No"),
            ("Physical inspection performed after maintenance?", "Yes,Partial,No"),
        ]),
        ("Equipment Removal", [
            ("Removal authorisation required?", "Yes,No"),
            ("Chain of custody documented?", "Yes,Partial,No"),
            ("Equipment inspected on return?", "Yes,No"),
            ("Return logged in asset management?", "Yes,No"),
        ]),
    ]

    row = 5
    col_headers = ["Assessment Question", "Status", "Evidence Reference", "Notes"]
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for section_name, questions in sections:
        ws.cell(row=row, column=1, value=section_name).font = Font(name="Calibri", bold=True)
        row += 1

        for question, options in questions:
            ws.cell(row=row, column=1, value=question).border = styles["border"]

            status_cell = ws.cell(row=row, column=2)
            status_cell.fill = styles["input_cell"]["fill"]
            status_cell.border = styles["border"]

            dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(status_cell)

            for col in [3, 4]:
                cell = ws.cell(row=row, column=col)
                cell.fill = styles["input_cell"]["fill"]
                cell.border = styles["border"]

            row += 1
        row += 1

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 40

# =============================================================================
# REMOTE MAINTENANCE SHEET
# =============================================================================

def create_remote_maintenance_sheet(ws, styles):
    """Create Remote Maintenance sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "Remote Maintenance\nPolicy Requirement: Remote maintenance should be authorised, secure, logged and monitored"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = [
        ("Access ID", 12),
        ("Access Type", 25),
        ("Provider/System", 25),
        ("Equipment Types", 25),
        ("Authorised", 18),
        ("Secure Connection", 18),
        ("Session Logged", 18),
        ("Session Monitored", 18),
        ("Disabled When Not Required", 22),
        ("Last Access Review", 18),
        ("Compliance Status", 18),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_access_type = DataValidation(
        type="list",
        formula1='"Vendor Remote Support,Internal Remote Management,Cloud Management Portal"',
        allow_blank=True
    )
    ws.add_data_validation(dv_access_type)

    dv_authorised = DataValidation(
        type="list",
        formula1='"Yes - Pre-approved,Yes - On-demand,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_authorised)

    dv_secure = DataValidation(
        type="list",
        formula1='"Yes - VPN,Yes - Encrypted,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_secure)

    dv_logged = DataValidation(
        type="list",
        formula1='"Yes - Automated,Yes - Manual,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_logged)

    dv_monitored = DataValidation(
        type="list",
        formula1='"Yes - Real-time,Yes - Recorded,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_monitored)

    dv_disabled = DataValidation(
        type="list",
        formula1='"Yes - Always,Yes - Usually,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_disabled)

    dv_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_access_type.add(ws.cell(row=r, column=2))
        dv_authorised.add(ws.cell(row=r, column=5))
        dv_secure.add(ws.cell(row=r, column=6))
        dv_logged.add(ws.cell(row=r, column=7))
        dv_monitored.add(ws.cell(row=r, column=8))
        dv_disabled.add(ws.cell(row=r, column=9))
        dv_status.add(ws.cell(row=r, column=11))

    ws.freeze_panes = "A4"

# =============================================================================
# SUMMARY DASHBOARD
# =============================================================================

def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "EQUIPMENT MAINTENANCE SUMMARY DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Period:"
    ws["A3"].font = Font(name="Calibri", bold=True)
    ws["B3"].fill = styles["input_cell"]["fill"]
    ws["B3"].border = styles["border"]

    ws["A5"] = "Compliance Metrics"
    ws["A5"].font = Font(name="Calibri", size=12, bold=True)

    headers = ["Domain", "Total Items", f"{CHECK} Compliant", f"{WARNING} Partial", f"{XMARK} Non-Compliant", "N/A", "% Compliant"]
    row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    domains = [
        ("Equipment Inventory", "Equipment Inventory", 13),
        ("Personnel & Vendors", "Personnel & Vendors", 11),
        ("Remote Maintenance", "Remote Maintenance", 11),
    ]

    row += 1
    start_data_row = row
    for label, sheet, status_col in domains:
        ws.cell(row=row, column=1, value=label)

        status_col_letter = get_column_letter(status_col)
        status_range = f"'{sheet}'!{status_col_letter}4:{status_col_letter}53"

        ws.cell(row=row, column=2, value=f'=COUNTA({status_range})')
        ws.cell(row=row, column=3, value=f'=COUNTIF({status_range},"{CHECK}*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({status_range},"{WARNING}*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({status_range},"{XMARK}*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({status_range},"N/A")')
        ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"N/A",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{row-1})")
        cell.font = Font(name="Calibri", bold=True)

    total_pct = ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"N/A",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
    total_pct.font = Font(name="Calibri", bold=True, color="0000FF", size=12)

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 18

# =============================================================================
# EVIDENCE REGISTER
# =============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Document all supporting evidence for audit traceability."
    ws["A2"].font = Font(name="Calibri", italic=True)

    headers = [
        ("Evidence ID", 12),
        ("Evidence Type", 18),
        ("Description", 40),
        ("Related Sheet/Item", 25),
        ("File Name", 30),
        ("File Location", 45),
        ("Collection Date", 15),
        ("Collected By", 20),
        ("Retention Period", 15),
        ("Notes", 40),
    ]

    row = 4
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Maintenance Record,Vendor Contract,Screenshot,Export,Log Sample,Procedure Document,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EVID-{r-4:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "A5"

# =============================================================================
# APPROVAL SIGN-OFF
# =============================================================================

def create_approval_signoff(ws, styles):
    """Create Approval and Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Summary"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Overall Compliance Rate", "='Summary Dashboard'!G10"),
        ("Assessment Status", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "" or value.startswith("="):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Approval Workflow"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    approval_headers = ["Role", "Name", "Signature", "Date", "Comments"]
    row += 3
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    roles = ["Assessor", "IT Operations Manager", "CISO", "Compliance Officer"]
    row += 1
    for role in roles:
        ws.cell(row=row, column=1, value=role).font = Font(name="Calibri", bold=True)
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 40

# =============================================================================
# MAIN WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    styles = setup_styles()

    ws = wb.create_sheet("Instructions & Legend", 0)
    create_instructions_sheet(ws, styles)

    ws = wb.create_sheet("Equipment Inventory")
    create_equipment_inventory_sheet(ws, styles)

    ws = wb.create_sheet("Maintenance Programme")
    create_maintenance_programme_sheet(ws, styles)

    ws = wb.create_sheet("Personnel & Vendors")
    create_personnel_vendors_sheet(ws, styles)

    ws = wb.create_sheet("Security During Maintenance")
    create_security_maintenance_sheet(ws, styles)

    ws = wb.create_sheet("Remote Maintenance")
    create_remote_maintenance_sheet(ws, styles)

    ws = wb.create_sheet("Summary Dashboard")
    create_summary_dashboard(ws, styles)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)

    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_signoff(ws, styles)

    return wb

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main entry point for workbook generation."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS Equipment Maintenance Assessment (A.7.13)")
        logger.info("=" * 70)

        wb = create_workbook()
        wb.save(OUTPUT_FILENAME)

        logger.info("%s SUCCESS: %s", CHECK, OUTPUT_FILENAME)
        logger.info("  %s 9 professional worksheets created", BULLET)
        logger.info("  %s Navy headers, yellow inputs styling applied", BULLET)
        logger.info("  %s Equipment inventory, personnel, remote access sheets", BULLET)
        logger.info("  %s Automated compliance dashboard with formulas", BULLET)
        logger.info("  %s Data validations and freeze panes configured", BULLET)
        logger.info("  %s Evidence register with audit traceability", BULLET)
        logger.info("  %s 4-level approval workflow", BULLET)
        logger.info("=" * 70)

        return 0

    except Exception as e:
        logger.error("%s ERROR: Failed to generate workbook", XMARK)
        logger.error("Error: %s", str(e))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
