#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.12-13.S3 - Maintenance Schedule Tracking Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.13: Equipment Maintenance
Assessment Domain 3 of 4: Maintenance Schedule Tracking

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates the Maintenance Schedule Tracking workbook for tracking
preventive maintenance schedules, monitoring compliance, and identifying
overdue items.

Reference Pattern: Based on ISMS-A.7.4-5-11 Physical Infrastructure Framework

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    pip3 install openpyxl

Usage:
    python3 generate_a712_3_maintenance_schedule.py

Output:
    ISMS-IMP-A.7.12-13.S3_Maintenance_Schedule_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.12-13.S3"
WORKBOOK_NAME = "Maintenance Schedule Tracking"
CONTROL_ID = "A.7.13"
CONTROL_NAME = "Equipment Maintenance"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = '\u2705'
XMARK = '\u274C'
WARNING = '\u26A0'
BULLET = '\u2022'

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_current": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_due_soon": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_overdue": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Configuration"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    config = [
        ("Organisation", ""),
        ("Workbook Start Date", ""),
        ("Responsible Department", ""),
        ("Update Frequency", "Monthly"),
        ("Days Before Due to Flag as Upcoming", "30"),
        ("Days Overdue to Escalate (Standard)", "14"),
        ("Days Overdue to Escalate (Critical)", "7"),
    ]

    row = 4
    for label, value in config:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "This is an OPERATIONAL TRACKING workbook - update regularly (monthly minimum).",
        "",
        "1. Sheet 2 (Equipment Schedule) - Master list of all equipment with maintenance schedules.",
        "2. Sheet 3 (Overdue Tracking) - Track overdue items with escalation and compensating controls.",
        "3. Sheet 4 (Upcoming Maintenance) - View maintenance due in next 30/60/90 days.",
        "4. Sheet 5 (Dashboard) - Automated compliance metrics and KPIs.",
        "5. Sheet 6 (Maintenance Log) - Historical record of completed maintenance.",
        "6. Sheet 7 (Evidence Register) - Link to supporting evidence.",
        "",
        "STATUS LEGEND:",
        f"  {CHECK} Current - Maintenance up to date",
        f"  {WARNING} Due Soon - Maintenance due within configured warning period",
        f"  {XMARK} Overdue - Maintenance past due date",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30

# =============================================================================
# EQUIPMENT SCHEDULE SHEET
# =============================================================================

def create_equipment_schedule_sheet(ws, styles):
    """Create Equipment Schedule sheet."""
    ws.merge_cells("A1:N1")
    ws["A1"] = "Equipment Maintenance Schedule\nTrack preventive maintenance for all equipment"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = [
        ("Equipment ID", 15),
        ("Equipment Type", 18),
        ("Equipment Description", 35),
        ("Location", 25),
        ("Criticality", 18),
        ("Maintenance Type", 22),
        ("Frequency", 15),
        ("Responsible Party", 20),
        ("Last Completed", 15),
        ("Next Due", 15),
        ("Status", 15),
        ("Days Until/Overdue", 18),
        ("Record Ref", 18),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_equip_type = DataValidation(
        type="list",
        formula1='"Server,Network Device,Storage,UPS,HVAC,Generator,Security System,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_equip_type)

    dv_criticality = DataValidation(
        type="list",
        formula1='"Tier 1 - Critical,Tier 2 - Standard"',
        allow_blank=True
    )
    ws.add_data_validation(dv_criticality)

    dv_maint_type = DataValidation(
        type="list",
        formula1='"Firmware Update,Inspection,Battery Check,Filter Replacement,Calibration,Full Service,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_maint_type)

    dv_frequency = DataValidation(
        type="list",
        formula1='"Monthly,Quarterly,Semi-annually,Annually"',
        allow_blank=True
    )
    ws.add_data_validation(dv_frequency)

    dv_responsible = DataValidation(
        type="list",
        formula1='"Internal - IT,Internal - Facilities,Vendor,Manufacturer"',
        allow_blank=True
    )
    ws.add_data_validation(dv_responsible)

    for r in range(4, 204):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_equip_type.add(ws.cell(row=r, column=2))
        dv_criticality.add(ws.cell(row=r, column=5))
        dv_maint_type.add(ws.cell(row=r, column=6))
        dv_frequency.add(ws.cell(row=r, column=7))
        dv_responsible.add(ws.cell(row=r, column=8))

        # Next Due formula (J column = I column + frequency days)
        ws.cell(row=r, column=10, value=(
            f'=IF(I{r}="","",I{r}+CHOOSE(MATCH(G{r},'
            '{"Monthly","Quarterly","Semi-annually","Annually"},0),30,90,180,365))'
        ))

        # Status formula (K column)
        ws.cell(row=r, column=11, value=(
            f'=IF(J{r}="","",IF(J{r}>TODAY()+30,"{CHECK} Current",'
            f'IF(J{r}>TODAY(),"{WARNING} Due Soon","{XMARK} Overdue")))'
        ))

        # Days Until/Overdue formula (L column)
        ws.cell(row=r, column=12, value=f'=IF(J{r}="","",J{r}-TODAY())')

    ws.freeze_panes = "A4"

# =============================================================================
# OVERDUE TRACKING SHEET
# =============================================================================

def create_overdue_tracking_sheet(ws, styles):
    """Create Overdue Tracking sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "Overdue Maintenance Tracking\nManage and escalate overdue maintenance items"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = [
        ("Equipment ID", 15),
        ("Equipment Description", 35),
        ("Maintenance Type", 22),
        ("Original Due Date", 18),
        ("Days Overdue", 15),
        ("Reason for Delay", 25),
        ("Estimated Completion", 18),
        ("Escalated", 12),
        ("Escalated To", 20),
        ("Compensating Control", 35),
        ("Resolution Notes", 40),
    ]

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_reason = DataValidation(
        type="list",
        formula1='"Parts on Order,Vendor Scheduling,Resource Unavailable,Budget Hold,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_reason)

    dv_yes_no = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_yes_no)

    for r in range(4, 54):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_reason.add(ws.cell(row=r, column=6))
        dv_yes_no.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"

# =============================================================================
# UPCOMING MAINTENANCE SHEET
# =============================================================================

def create_upcoming_maintenance_sheet(ws, styles):
    """Create Upcoming Maintenance sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Upcoming Maintenance\nPlan maintenance for the next 30/60/90 days"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # Instructions
    ws["A3"] = "This sheet provides a view of upcoming maintenance. Data is linked from Equipment Schedule."
    ws["A3"].font = Font(name="Calibri", italic=True)

    ws["A5"] = "Next 30 Days"
    ws["A5"].font = Font(name="Calibri", size=12, bold=True)
    ws["A5"].fill = styles["status_due_soon"]["fill"]

    columns = [
        ("Equipment ID", 15),
        ("Equipment Description", 35),
        ("Maintenance Type", 22),
        ("Due Date", 15),
        ("Days Until Due", 15),
        ("Scheduled Date", 15),
        ("Assigned To", 20),
        ("Vendor Booked", 15),
    ]

    row = 6
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_yes_no = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_yes_no)

    for r in range(7, 37):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_yes_no.add(ws.cell(row=r, column=8))

    # 31-60 Days section
    row = 39
    ws[f"A{row}"] = "31-60 Days"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)

    row += 1
    for col_idx, (col_name, _) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    for r in range(row + 1, row + 21):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # 61-90 Days section
    row = 62
    ws[f"A{row}"] = "61-90 Days"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)

    row += 1
    for col_idx, (col_name, _) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    for r in range(row + 1, row + 21):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

# =============================================================================
# DASHBOARD SHEET
# =============================================================================

def create_dashboard_sheet(ws, styles):
    """Create Dashboard sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "MAINTENANCE SCHEDULE DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Overall Compliance"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    metrics = [
        ("Total Equipment in Programme", f"=COUNTA('Equipment Schedule'!A4:A203)"),
        ("Equipment Current", f"=COUNTIF('Equipment Schedule'!K4:K203,\"{CHECK}*\")"),
        ("Equipment Due Soon", f"=COUNTIF('Equipment Schedule'!K4:K203,\"{WARNING}*\")"),
        ("Equipment Overdue", f"=COUNTIF('Equipment Schedule'!K4:K203,\"{XMARK}*\")"),
        ("Compliance Rate (%)", f"=IF(B4=0,\"N/A\",ROUND(B5/B4*100,1)&\"%\")"),
    ]

    row = 4
    for label, formula in metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "Compliance by Criticality"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    row += 2
    headers = ["Criticality", "Total", "Current", "Due Soon", "Overdue", "% Current"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    ws.cell(row=row, column=1, value="Tier 1 - Critical")
    ws.cell(row=row, column=2, value="=COUNTIF('Equipment Schedule'!E4:E203,\"Tier 1*\")")
    ws.cell(row=row, column=3, value=f"=COUNTIFS('Equipment Schedule'!E4:E203,\"Tier 1*\",'Equipment Schedule'!K4:K203,\"{CHECK}*\")")
    ws.cell(row=row, column=4, value=f"=COUNTIFS('Equipment Schedule'!E4:E203,\"Tier 1*\",'Equipment Schedule'!K4:K203,\"{WARNING}*\")")
    ws.cell(row=row, column=5, value=f"=COUNTIFS('Equipment Schedule'!E4:E203,\"Tier 1*\",'Equipment Schedule'!K4:K203,\"{XMARK}*\")")
    ws.cell(row=row, column=6, value=f"=IF(B{row}=0,\"N/A\",ROUND(C{row}/B{row}*100,1)&\"%\")")

    row += 1
    ws.cell(row=row, column=1, value="Tier 2 - Standard")
    ws.cell(row=row, column=2, value="=COUNTIF('Equipment Schedule'!E4:E203,\"Tier 2*\")")
    ws.cell(row=row, column=3, value=f"=COUNTIFS('Equipment Schedule'!E4:E203,\"Tier 2*\",'Equipment Schedule'!K4:K203,\"{CHECK}*\")")
    ws.cell(row=row, column=4, value=f"=COUNTIFS('Equipment Schedule'!E4:E203,\"Tier 2*\",'Equipment Schedule'!K4:K203,\"{WARNING}*\")")
    ws.cell(row=row, column=5, value=f"=COUNTIFS('Equipment Schedule'!E4:E203,\"Tier 2*\",'Equipment Schedule'!K4:K203,\"{XMARK}*\")")
    ws.cell(row=row, column=6, value=f"=IF(B{row}=0,\"N/A\",ROUND(C{row}/B{row}*100,1)&\"%\")")

    ws[f"A{row+3}"] = "Overdue Summary"
    ws[f"A{row+3}"].font = Font(name="Calibri", size=12, bold=True)

    row += 4
    ws.cell(row=row, column=1, value="Total Overdue Items")
    ws.cell(row=row, column=2, value=f"=COUNTIF('Equipment Schedule'!K4:K203,\"{XMARK}*\")")
    row += 1
    ws.cell(row=row, column=1, value="Critical Equipment Overdue")
    ws.cell(row=row, column=2, value=f"=COUNTIFS('Equipment Schedule'!E4:E203,\"Tier 1*\",'Equipment Schedule'!K4:K203,\"{XMARK}*\")")

    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 20

# =============================================================================
# MAINTENANCE LOG SHEET
# =============================================================================

def create_maintenance_log_sheet(ws, styles):
    """Create Maintenance Log sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Maintenance Log\nHistorical record of completed maintenance"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = [
        ("Equipment ID", 15),
        ("Maintenance Type", 22),
        ("Scheduled Date", 15),
        ("Actual Completion", 18),
        ("Performed By", 25),
        ("Record Reference", 20),
        ("Findings/Notes", 40),
        ("Follow-up Required", 18),
        ("Follow-up Notes", 40),
    ]

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_yes_no = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_yes_no)

    for r in range(4, 504):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_yes_no.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"

# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================

def create_evidence_register_sheet(ws, styles):
    """Create Evidence Register sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = [
        ("Evidence ID", 12),
        ("Evidence Type", 20),
        ("Description", 40),
        ("Related Equipment/Item", 25),
        ("File Name", 30),
        ("File Location", 45),
        ("Collection Date", 15),
        ("Collected By", 20),
        ("Retention Period", 15),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Maintenance Ticket,Vendor Report,Test Report,Inspection Certificate,Log Export,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"EVID-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "A4"

# =============================================================================
# MAIN WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    styles = setup_styles()

    ws = wb.create_sheet("Instructions", 0)
    create_instructions_sheet(ws, styles)

    ws = wb.create_sheet("Equipment Schedule")
    create_equipment_schedule_sheet(ws, styles)

    ws = wb.create_sheet("Overdue Tracking")
    create_overdue_tracking_sheet(ws, styles)

    ws = wb.create_sheet("Upcoming Maintenance")
    create_upcoming_maintenance_sheet(ws, styles)

    ws = wb.create_sheet("Dashboard")
    create_dashboard_sheet(ws, styles)

    ws = wb.create_sheet("Maintenance Log")
    create_maintenance_log_sheet(ws, styles)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register_sheet(ws, styles)

    return wb

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main entry point for workbook generation."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS Maintenance Schedule Tracking (A.7.13)")
        logger.info("=" * 70)

        wb = create_workbook()
        wb.save(OUTPUT_FILENAME)

        logger.info("%s SUCCESS: %s", CHECK, OUTPUT_FILENAME)
        logger.info("  %s 7 operational tracking worksheets created", BULLET)
        logger.info("  %s 200 equipment rows in Equipment Schedule", BULLET)
        logger.info("  %s Automated status formulas (Current/Due Soon/Overdue)", BULLET)
        logger.info("  %s Dashboard with compliance metrics", BULLET)
        logger.info("  %s Overdue tracking with escalation", BULLET)
        logger.info("  %s Upcoming maintenance planning (30/60/90 days)", BULLET)
        logger.info("  %s 500 row Maintenance Log for history", BULLET)
        logger.info("=" * 70)

        return 0

    except Exception as e:
        logger.error("%s ERROR: Failed to generate workbook", XMARK)
        logger.error("Error: %s", str(e))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
