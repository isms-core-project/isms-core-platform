#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS Assessment File Normalization Utility - A.7.12-13 Infrastructure Maintenance
================================================================================

ISO/IEC 27001:2022 Controls A.7.12-13: Cabling Security and Equipment Maintenance
Quality Assurance Utility: Excel Assessment File Normalization & Validation

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script normalizes A.7.12-13 assessment workbooks for dashboard integration by:
1. Scanning for completed assessment files
2. Validating document IDs in Instructions sheet
3. Copying to normalized filenames (no dates/versions)
4. Creating audit manifest for traceability

Reference Pattern: Based on ISMS-A.7.4 Physical Security Monitoring Framework

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel processing

Installation:
    pip3 install openpyxl

Usage:
    python3 normalize_assessment_files_a712.py

Output:
    Normalized assessment workbooks + validation report

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import os
import sys
import shutil
from datetime import datetime
from pathlib import Path

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    logger.error("openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)

# =============================================================================
# CONFIGURATION
# =============================================================================

EXPECTED_DOCS = {
    "ISMS-IMP-A.7.12-13.S1": {
        "title": "Cabling Security Assessment",
        "normalized": "ISMS-IMP-A.7.12-13.S1_Cabling_Security.xlsx"
    },
    "ISMS-IMP-A.7.12-13.S2": {
        "title": "Equipment Maintenance Assessment",
        "normalized": "ISMS-IMP-A.7.12-13.S2_Equipment_Maintenance.xlsx"
    },
    "ISMS-IMP-A.7.12-13.S3": {
        "title": "Maintenance Schedule Tracking",
        "normalized": "ISMS-IMP-A.7.12-13.S3_Maintenance_Schedule.xlsx"
    },
    "ISMS-IMP-A.7.12-13.S4": {
        "title": "Infrastructure Compliance Dashboard",
        "normalized": "ISMS-IMP-A.7.12-13.S4_Compliance_Dashboard.xlsx"
    },
}

# =============================================================================
# VALIDATION FUNCTIONS
# =============================================================================

def validate_workbook(filepath):
    """
    Validate workbook contains expected document ID in Instructions sheet.

    Args:
        filepath: Path to Excel workbook

    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        # Check for Instructions sheet (various naming conventions)
        instructions_sheet = None
        for sheet_name in ["Instructions & Legend", "Instructions", "Instructions & Config"]:
            if sheet_name in wb.sheetnames:
                instructions_sheet = wb[sheet_name]
                break

        if instructions_sheet is None:
            wb.close()
            return (None, None)

        # Look for Document ID in first 25 rows
        for row in range(1, 25):
            cell_label = instructions_sheet.cell(row=row, column=1).value

            if cell_label and "Document ID" in str(cell_label):
                doc_id_value = instructions_sheet.cell(row=row, column=2).value

                if doc_id_value:
                    doc_id = str(doc_id_value).strip()

                    if doc_id in EXPECTED_DOCS:
                        title = EXPECTED_DOCS[doc_id]["title"]
                        wb.close()
                        return (doc_id, title)

        wb.close()
        return (None, None)

    except Exception as e:
        logger.warning("Error reading file %s: %s", filepath, e)
        return (None, None)


def get_file_info(filepath):
    """Get file metadata for manifest."""
    stat = filepath.stat()
    return {
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime),
        "created": datetime.fromtimestamp(stat.st_ctime),
    }

# =============================================================================
# SCANNING FUNCTIONS
# =============================================================================

def scan_directory(directory):
    """
    Scan directory for A.7.12-13 assessment workbooks.

    Args:
        directory: Path to scan

    Returns:
        dict: {doc_id: {'path': Path, 'info': metadata}}
    """
    found_assessments = {}
    directory = Path(directory)

    xlsx_files = [f for f in directory.glob("*.xlsx") if not f.name.startswith("~$")]

    if not xlsx_files:
        logger.warning("No Excel files found in %s", directory)
        return found_assessments

    logger.info("Scanning %d Excel file(s) in %s...", len(xlsx_files), directory)

    for filepath in sorted(xlsx_files):
        logger.info("  Checking: %s", filepath.name)
        doc_id, title = validate_workbook(filepath)

        if doc_id:
            logger.info("    Valid: %s - %s", doc_id, title)

            if doc_id in found_assessments:
                logger.warning("    DUPLICATE FOUND FOR %s", doc_id)
                logger.warning("    Previous: %s", found_assessments[doc_id]['path'].name)
                logger.warning("    Current: %s", filepath.name)
                # Keep the newer file
                if filepath.stat().st_mtime > found_assessments[doc_id]['path'].stat().st_mtime:
                    found_assessments[doc_id] = {
                        'path': filepath,
                        'info': get_file_info(filepath)
                    }
                    logger.info("    Using newer file")
            else:
                found_assessments[doc_id] = {
                    'path': filepath,
                    'info': get_file_info(filepath)
                }
        else:
            logger.info("    Skipped (not a valid A.7.12-13 assessment)")

    return found_assessments

# =============================================================================
# MANIFEST CREATION
# =============================================================================

def create_manifest(output_dir, mapping, source_dir):
    """
    Create audit manifest documenting the normalization process.

    Args:
        output_dir: Destination directory
        mapping: Dictionary of normalized files
        source_dir: Source directory

    Returns:
        Path: Path to created manifest file
    """
    manifest_path = Path(output_dir) / "normalization_manifest_a712.txt"

    with open(manifest_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("ISMS ASSESSMENT FILE NORMALIZATION MANIFEST\n")
        f.write("ISO/IEC 27001:2022 - Controls A.7.12-13: Cabling Security & Equipment Maintenance\n")
        f.write("=" * 80 + "\n\n")

        f.write("NORMALIZATION METADATA\n")
        f.write("-" * 80 + "\n")
        f.write(f"Normalization Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Source Directory:        {source_dir.resolve()}\n")
        f.write(f"Output Directory:        {output_dir.resolve()}\n")
        f.write(f"Files Normalized:        {len(mapping)}/4 required\n")
        f.write(f"Normalization Status:    {'COMPLETE' if len(mapping) == 4 else 'INCOMPLETE'}\n")
        f.write("\n")

        f.write("FILE MAPPING\n")
        f.write("-" * 80 + "\n\n")

        for doc_id in sorted(EXPECTED_DOCS.keys()):
            if doc_id in mapping:
                info = mapping[doc_id]
                source_path = info['path']
                normalized_name = info['normalized']
                file_info = info['info']

                f.write(f"Document ID:     {doc_id}\n")
                f.write(f"Document Title:  {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"Source File:     {source_path.name}\n")
                f.write(f"Normalized File: {normalized_name}\n")
                f.write(f"File Size:       {file_info['size']:,} bytes\n")
                f.write(f"Last Modified:   {file_info['modified'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("Status:          NORMALIZED\n")
                f.write("\n")
            else:
                f.write(f"Document ID:     {doc_id}\n")
                f.write(f"Document Title:  {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"Normalized File: {EXPECTED_DOCS[doc_id]['normalized']}\n")
                f.write("Status:          NOT FOUND\n")
                f.write("\n")

        f.write("=" * 80 + "\n")
        f.write("EXPECTED ASSESSMENTS\n")
        f.write("=" * 80 + "\n\n")

        f.write("This control requires 4 assessment workbooks:\n\n")

        for doc_id, info in EXPECTED_DOCS.items():
            status = "OK" if doc_id in mapping else "MISSING"
            f.write(f"  [{status}] {doc_id}: {info['title']}\n")
            f.write(f"         Normalized: {info['normalized']}\n\n")

        f.write("=" * 80 + "\n")
        f.write("END OF MANIFEST\n")
        f.write("=" * 80 + "\n")

    return manifest_path

# =============================================================================
# MAIN NORMALIZATION FUNCTION
# =============================================================================

def normalize_files(source_dir=None, output_dir=None, auto_confirm=False):
    """
    Main normalization orchestration function.

    Args:
        source_dir: Source directory (prompts if None)
        output_dir: Output directory (prompts if None)
        auto_confirm: Skip confirmation prompts (for automation)

    Returns:
        bool: True if successful, False otherwise
    """
    logger.info("=" * 70)
    logger.info("ISMS Assessment File Normalization - A.7.12-13")
    logger.info("Controls: Cabling Security and Equipment Maintenance")
    logger.info("=" * 70)

    if not source_dir:
        source_dir = input("Enter source directory (or press Enter for current): ").strip()
        if not source_dir:
            source_dir = "."

    source_dir = Path(source_dir).resolve()

    if not source_dir.exists() or not source_dir.is_dir():
        logger.error("Source directory does not exist: %s", source_dir)
        return False

    logger.info("Source directory: %s", source_dir)

    if not output_dir:
        default_output = source_dir / "Dashboard_Sources"
        output_input = input(f"Enter output directory (or press Enter for '{default_output.name}'): ").strip()
        if not output_input:
            output_dir = default_output
        else:
            output_dir = Path(output_input)

    output_dir = Path(output_dir).resolve()

    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        logger.info("Output directory: %s", output_dir)
    except Exception as e:
        logger.error("Error creating output directory: %s", e)
        return False

    found = scan_directory(source_dir)

    if not found:
        logger.error("No valid A.7.12-13 assessment workbooks found")
        return False

    logger.info("Found %d of %d required assessments", len(found), len(EXPECTED_DOCS))

    mapping = {}
    for doc_id in sorted(EXPECTED_DOCS.keys()):
        if doc_id in found:
            normalized_name = EXPECTED_DOCS[doc_id]["normalized"]
            source_path = found[doc_id]['path']

            logger.info("  %s: %s -> %s", doc_id, source_path.name, normalized_name)

            mapping[doc_id] = {
                'path': source_path,
                'normalized': normalized_name,
                'info': found[doc_id]['info']
            }
        else:
            logger.warning("  %s: NOT FOUND", doc_id)

    if not auto_confirm:
        response = input("Proceed with normalization? (y/n): ").strip().lower()
        if response != 'y':
            logger.info("Normalization cancelled by user")
            return False

    logger.info("Normalizing files...")

    for doc_id, info in mapping.items():
        source = info['path']
        dest = output_dir / info['normalized']

        logger.info("  Copying: %s -> %s", source.name, dest.name)

        try:
            shutil.copy2(source, dest)
        except Exception as e:
            logger.error("Error copying %s: %s", source.name, e)
            return False

    logger.info("Creating audit manifest...")
    try:
        manifest = create_manifest(output_dir, mapping, source_dir)
        logger.info("  Created: %s", manifest.name)
    except Exception as e:
        logger.error("Error creating manifest: %s", e)
        return False

    logger.info("=" * 70)
    logger.info("NORMALIZATION COMPLETE")
    logger.info("=" * 70)
    logger.info("Normalized files: %s", output_dir)
    logger.info("Audit manifest:   %s", manifest)

    return True

# =============================================================================
# COMMAND LINE INTERFACE
# =============================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Normalize A.7.12-13 assessment workbooks for dashboard integration"
    )

    parser.add_argument(
        '--source', '-s',
        help='Source directory containing assessment workbooks'
    )

    parser.add_argument(
        '--output', '-o',
        help='Output directory for normalized files'
    )

    parser.add_argument(
        '--auto-confirm', '-y',
        action='store_true',
        help='Skip confirmation prompts'
    )

    args = parser.parse_args()

    success = normalize_files(
        source_dir=args.source,
        output_dir=args.output,
        auto_confirm=args.auto_confirm
    )

    sys.exit(0 if success else 1)

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED (syntax validated, structure verified)
# QA_TOOL: Claude Code Deep Scan
# QA_NOTE: STANDARDIZATION - License header, logging, main() pattern applied
# =============================================================================
