#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS CORE Privacy & Cloud — Compliance Checklist Engine
================================================================================

Shared engine for all Privacy (ISO/IEC 27701:2025) and Cloud (ISO/IEC 27018:2025)
compliance checklist generators.  Each control group has a thin generator that
defines REQUIREMENTS and calls this engine.  All workbook structure, styling,
formulas, and sheet creation logic lives here — written once, used by 33
generators (21 Privacy + 12 Cloud).

Workbook structure (identical for every control group):
  1. Executive Summary — board-level overview, traffic light, approval
  2. Dashboard — formula-driven aggregation from domain sheets
  3–N. Domain sheets — one per requirements domain (checklist rows)

Design principles:
  - Every "shall" statement in a PRIV-POL or CLD-POL = one checklist row
  - Status dropdown: Compliant / Partial / Non-Compliant / N/A
  - Traffic light scoring: GREEN >= 90% / AMBER >= 70% / RED < 70%
  - Dashboard uses COUNTIF formulas referencing domain sheets
  - All formulas are dynamic (adapt to any number of domains/requirements)

Key difference from op_checklist_engine.py:
  - generate_checklist() accepts iso_standard parameter (default ISO/IEC 27701:2025)
  - Source policy derived from document_id by replacing -CHK- with -POL-

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# =============================================================================
# LOGGING
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# COLOUR CONSTANTS (Framework palette)
# =============================================================================
C_NAVY = "003366"
C_BLUE = "4472C4"
C_GRAY = "D9D9D9"
C_LGRAY = "F2F2F2"
C_YELLOW = "FFFFCC"
C_GREEN = "C6EFCE"
C_AMBER = "FFEB9C"
C_RED = "FFC7CE"
C_WHITE = "FFFFFF"


# =============================================================================
# STYLE SYSTEM
# =============================================================================

def setup_styles():
    """Define all cell styles (Framework palette)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color=C_WHITE),
            "fill": PatternFill(start_color=C_NAVY, end_color=C_NAVY, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color=C_WHITE),
            "fill": PatternFill(start_color=C_BLUE, end_color=C_BLUE, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "section_title": {
            "font": Font(name="Calibri", size=12, bold=True, color=C_NAVY),
            "alignment": Alignment(horizontal="left", vertical="center"),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color=C_GRAY, end_color=C_GRAY, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color=C_YELLOW, end_color=C_YELLOW, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "formula_cell": {
            "font": Font(name="Calibri", size=10, color=C_BLUE, italic=True),
            "fill": PatternFill(start_color=C_LGRAY, end_color=C_LGRAY, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "total_row": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color=C_NAVY, end_color=C_NAVY, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
    }
    return styles


def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    for attr in ("font", "fill", "alignment", "border", "number_format"):
        if attr in style_dict:
            setattr(cell, attr, style_dict[attr])


def set_column_widths(ws, widths):
    """Set column widths from dict."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def write_banner(ws, styles, title, row=1, end_col="H"):
    """Write merged navy header banner."""
    ws.merge_cells(f"A{row}:{end_col}{row}")
    cell = ws[f"A{row}"]
    cell.value = title
    apply_style(cell, styles["header"])
    ws.row_dimensions[row].height = 40


# =============================================================================
# SHEET CREATORS
# =============================================================================

def create_executive_summary(ws, styles, domain_config, document_id, control_name, source_policy):
    """Create Executive Summary sheet.

    Args:
        ws: worksheet
        styles: style dict from setup_styles()
        domain_config: list of (domain_name, start_row, end_row) tuples
        document_id: e.g. "PRIV-CHK-A.1.2.2-5"
        control_name: e.g. "Lawful Basis and Consent"
        source_policy: e.g. "PRIV-POL-A.1.2.2-5"
    """
    logger.info("Creating Executive Summary sheet...")

    set_column_widths(ws, {"A": 28, "B": 22, "C": 22, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18})
    ws.sheet_properties.tabColor = C_NAVY

    total_row_dash = 4 + len(domain_config)  # Dashboard total row

    # --- Title ---
    write_banner(ws, styles, f"{document_id} \u2014 {control_name} Compliance Checklist")

    # --- Document Info ---
    row = 3
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    apply_style(ws[f"A{row}"], styles["section_title"])
    row += 1

    doc_info = [
        ("Document ID", document_id),
        ("Source Policy", source_policy),
        ("ISO Controls", control_name),
        ("Generated", datetime.now().strftime("%d.%m.%Y")),
        ("Prepared By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Annual"),
    ]
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = value
        if not value:
            apply_style(ws[f"B{row}"], styles["input_cell"])
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # --- Overall Compliance ---
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "OVERALL COMPLIANCE STATUS"
    apply_style(ws[f"A{row}"], styles["subheader"])
    row += 1

    ws[f"A{row}"] = "Overall Compliance"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)
    ws[f"B{row}"] = f"='Dashboard'!G{total_row_dash}"
    ws[f"B{row}"].font = Font(name="Calibri", size=20, bold=True)
    ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    ws[f"A{row}"] = "Status"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)
    ws[f"B{row}"] = (
        f'=IF(Dashboard!G{total_row_dash}="N/A","NOT ASSESSED",'
        f'IF(ISNUMBER(VALUE(SUBSTITUTE(Dashboard!G{total_row_dash},"%",""))),'
        f'IF(VALUE(SUBSTITUTE(Dashboard!G{total_row_dash},"%",""))>=90,"GREEN",'
        f'IF(VALUE(SUBSTITUTE(Dashboard!G{total_row_dash},"%",""))>=70,"AMBER","RED")),'
        f'"NOT ASSESSED"))'
    )
    ws[f"B{row}"].font = Font(name="Calibri", size=16, bold=True)
    ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    # --- Key Metrics ---
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "KEY METRICS"
    apply_style(ws[f"A{row}"], styles["subheader"])
    row += 1

    # Build owner coverage formula dynamically
    owner_parts = []
    for dname, dstart, dend in domain_config:
        owner_parts.append(f"COUNTA('{dname}'!F{dstart}:F{dend})")
    owner_formula = (
        f'=IF(Dashboard!B{total_row_dash}=0,"N/A",'
        f'ROUND(({"+".join(owner_parts)})'
        f'/Dashboard!B{total_row_dash}*100,0)&"%")'
    )

    metrics = [
        ("Total Requirements", f"='Dashboard'!B{total_row_dash}"),
        ("Compliant", f"='Dashboard'!C{total_row_dash}"),
        ("Partial", f"='Dashboard'!D{total_row_dash}"),
        ("Non-Compliant", f"='Dashboard'!E{total_row_dash}"),
        ("Not Applicable", f"='Dashboard'!F{total_row_dash}"),
        ("Owner Coverage", owner_formula),
    ]
    for label, formula in metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = formula
        apply_style(ws[f"B{row}"], styles["formula_cell"])
        ws[f"B{row}"].font = Font(name="Calibri", size=14, bold=True, color=C_BLUE)
        row += 1

    # --- Domain Status ---
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "DOMAIN STATUS SUMMARY"
    apply_style(ws[f"A{row}"], styles["subheader"])
    row += 1

    for ci, h in enumerate(["Domain", "Compliance %", "Compliant", "Gaps", "Status"], 1):
        cell = ws.cell(row=row, column=ci, value=h)
        apply_style(cell, styles["column_header"])
    row += 1

    for di, (dname, dstart, dend) in enumerate(domain_config):
        dash_row = 4 + di
        ws.cell(row=row, column=1, value=dname).border = styles["border"]
        ws.cell(row=row, column=2, value=f"='Dashboard'!G{dash_row}").border = styles["border"]
        ws.cell(row=row, column=3, value=f"='Dashboard'!C{dash_row}").border = styles["border"]
        ws.cell(row=row, column=4, value=f"='Dashboard'!E{dash_row}").border = styles["border"]
        ws.cell(row=row, column=5,
                value=f'=IF(Dashboard!G{dash_row}="N/A","--",'
                      f'IF(ISNUMBER(VALUE(SUBSTITUTE(Dashboard!G{dash_row},"%",""))),'
                      f'IF(VALUE(SUBSTITUTE(Dashboard!G{dash_row},"%",""))>=90,"PASS",'
                      f'IF(VALUE(SUBSTITUTE(Dashboard!G{dash_row},"%",""))>=70,"REVIEW","FAIL")),'
                      f'"--"))').border = styles["border"]
        row += 1

    # --- Critical Issues ---
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "CRITICAL ISSUES"
    apply_style(ws[f"A{row}"], styles["subheader"])
    row += 1

    for ci, h in enumerate(["#", "Issue Description", "Domain", "Owner", "Due Date"], 1):
        cell = ws.cell(row=row, column=ci, value=h)
        apply_style(cell, styles["column_header"])
    row += 1

    for i in range(1, 6):
        ws.cell(row=row, column=1, value=i).border = styles["border"]
        for col in range(2, 6):
            apply_style(ws.cell(row=row, column=col), styles["input_cell"])
        row += 1

    # --- Management Commentary ---
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "MANAGEMENT COMMENTARY"
    apply_style(ws[f"A{row}"], styles["subheader"])
    row += 1

    ws.merge_cells(f"A{row}:H{row + 2}")
    cell = ws[f"A{row}"]
    cell.value = f"[Enter management commentary on {control_name.lower()} posture]"
    apply_style(cell, styles["input_cell"])
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 60
    row += 3

    # --- Approval ---
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "APPROVAL"
    apply_style(ws[f"A{row}"], styles["subheader"])
    row += 1

    dv_decision = DataValidation(
        type="list", formula1='"Approved,Approved with Conditions,Not Approved"', allow_blank=True)
    ws.add_data_validation(dv_decision)

    for role in ["Assessed By", "Reviewed By (DPO)", "Approved By (CEO / CISO)"]:
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = role
        ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
        row += 1
        for field in ["Name:", "Date:", "Signature:", "Decision:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", size=10)
            ws[f"A{row}"].border = styles["border"]
            cell = ws[f"B{row}"]
            apply_style(cell, styles["input_cell"])
            if field == "Decision:":
                dv_decision.add(cell)
            row += 1
        row += 1

    # Print setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.print_area = f"A1:H{row}"

    logger.info(f"  Executive Summary complete ({row} rows)")


def create_dashboard(ws, styles, domain_config, control_name):
    """Create Dashboard sheet — formula-driven aggregation.

    Args:
        ws: worksheet
        styles: style dict
        domain_config: list of (domain_name, start_row, end_row) tuples
        control_name: e.g. "Lawful Basis and Consent"
    """
    logger.info("Creating Dashboard sheet...")

    set_column_widths(ws, {"A": 28, "B": 14, "C": 14, "D": 14, "E": 16, "F": 10, "G": 16})
    ws.sheet_properties.tabColor = C_BLUE

    write_banner(ws, styles, f"{control_name} \u2014 Compliance Dashboard", end_col="G")

    # --- Domain Summary ---
    row = 3
    for ci, h in enumerate(["Domain", "Total Reqs", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"], 1):
        cell = ws.cell(row=row, column=ci, value=h)
        apply_style(cell, styles["column_header"])

    row = 4
    for dname, dstart, dend in domain_config:
        sr = f"'{dname}'!D{dstart}:D{dend}"
        ws.cell(row=row, column=1, value=dname)
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=row, column=1).border = styles["border"]
        ws.cell(row=row, column=2, value=f"=COUNTA({sr})")
        ws.cell(row=row, column=3, value=f'=COUNTIF({sr},"Compliant")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({sr},"Partial")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({sr},"Non-Compliant")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({sr},"N/A")')
        ws.cell(row=row, column=7,
                value=f'=IF((B{row}-F{row})=0,"N/A",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        for col in range(2, 8):
            apply_style(ws.cell(row=row, column=col), styles["formula_cell"])
        row += 1

    # TOTAL row
    total_row = row
    ws.cell(row=row, column=1, value="TOTAL")
    apply_style(ws.cell(row=row, column=1), styles["total_row"])
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=11, bold=True, color=C_WHITE)
    for col in range(2, 7):
        letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=SUM({letter}4:{letter}{row - 1})")
        apply_style(ws.cell(row=row, column=col), styles["total_row"])
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=11, bold=True, color=C_WHITE)
    ws.cell(row=row, column=7,
            value=f'=IF((B{row}-F{row})=0,"N/A",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
    apply_style(ws.cell(row=row, column=7), styles["total_row"])
    ws.cell(row=row, column=7).font = Font(name="Calibri", size=11, bold=True, color=C_WHITE)

    # --- Status Breakdown ---
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "COMPLIANCE BY STATUS"
    apply_style(ws[f"A{row}"], styles["subheader"])
    row += 1

    for label, formula, colour in [
        ("Compliant", f"=C{total_row}", C_GREEN),
        ("Partial", f"=D{total_row}", C_AMBER),
        ("Non-Compliant", f"=E{total_row}", C_RED),
        ("N/A", f"=F{total_row}", C_LGRAY),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=row, column=1).border = styles["border"]
        cell = ws.cell(row=row, column=2, value=formula)
        cell.font = Font(name="Calibri", size=14, bold=True)
        cell.fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = styles["border"]
        ws.cell(row=row, column=3,
                value=f'=IF(B{total_row}=0,"",ROUND(B{row}/B{total_row}*100,1)&"%")')
        ws.cell(row=row, column=3).border = styles["border"]
        row += 1

    # --- Top Gaps ---
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "TOP GAPS (Manual Entry)"
    apply_style(ws[f"A{row}"], styles["subheader"])
    row += 1

    for ci, h in enumerate(["Priority", "Domain", "Requirement", "Owner", "Due Date", "Status"], 1):
        cell = ws.cell(row=row, column=ci, value=h)
        apply_style(cell, styles["column_header"])
    row += 1

    dv_gap = DataValidation(type="list", formula1='"Open,In Progress,Closed,Deferred"', allow_blank=True)
    ws.add_data_validation(dv_gap)

    for _ in range(10):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, styles["input_cell"])
            if col == 6:
                dv_gap.add(cell)
        row += 1

    # --- Trend ---
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "COMPLIANCE TREND (Manual Entry)"
    apply_style(ws[f"A{row}"], styles["subheader"])
    row += 1

    for ci, h in enumerate(["Assessment Date", "Overall %", "Compliant", "Gaps", "Assessor", "Notes"], 1):
        cell = ws.cell(row=row, column=ci, value=h)
        apply_style(cell, styles["column_header"])
    row += 1

    for _ in range(4):
        for col in range(1, 7):
            apply_style(ws.cell(row=row, column=col), styles["input_cell"])
        row += 1

    logger.info(f"  Dashboard complete ({row} rows)")


def create_domain_sheet(ws, styles, domain_name, requirements, control_ref):
    """Create a domain checklist sheet.

    Args:
        ws: worksheet
        styles: style dict
        domain_name: sheet/domain title
        requirements: list of (req_id, section_label, requirement_text) tuples
        control_ref: ISO reference string
    Returns:
        last_data_row (int)
    """
    logger.info(f"Creating domain sheet: {domain_name} ({len(requirements)} requirements)...")

    set_column_widths(ws, {
        "A": 12, "B": 26, "C": 56, "D": 17, "E": 36, "F": 19, "G": 15, "H": 32
    })
    ws.sheet_properties.tabColor = C_BLUE

    write_banner(ws, styles, f"{domain_name} \u2014 Compliance Checklist")

    ws.merge_cells("A2:H2")
    ws["A2"] = control_ref
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=C_BLUE)
    ws["A2"].alignment = Alignment(horizontal="center")

    col_headers = ["Req #", "Section", "Requirement", "Status", "Evidence", "Owner", "Due Date", "Notes"]
    for ci, h in enumerate(col_headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        apply_style(cell, styles["column_header"])

    dv_status = DataValidation(
        type="list", formula1='"Compliant,Partial,Non-Compliant,N/A"', allow_blank=True)
    dv_status.error = "Select: Compliant, Partial, Non-Compliant, or N/A"
    ws.add_data_validation(dv_status)

    row = 4
    for req_id, section, text in requirements:
        ws.cell(row=row, column=1, value=req_id)
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=9, color="666666")
        ws.cell(row=row, column=1).border = styles["border"]

        ws.cell(row=row, column=2, value=section)
        ws.cell(row=row, column=2).font = Font(name="Calibri", size=9)
        ws.cell(row=row, column=2).border = styles["border"]
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", wrap_text=True)

        ws.cell(row=row, column=3, value=text)
        apply_style(ws.cell(row=row, column=3), styles["data_cell"])

        status_cell = ws.cell(row=row, column=4)
        apply_style(status_cell, styles["input_cell"])
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        dv_status.add(status_cell)

        for col in range(5, 9):
            apply_style(ws.cell(row=row, column=col), styles["input_cell"])

        ws.cell(row=row, column=7).number_format = "DD.MM.YYYY"
        ws.row_dimensions[row].height = 40
        row += 1

    last_data_row = row - 1
    _apply_conditional_formatting(ws, "D", 4, last_data_row)
    ws.freeze_panes = "A4"

    logger.info(f"  {domain_name} complete ({len(requirements)} requirements, rows 4-{last_data_row})")
    return last_data_row


def _apply_conditional_formatting(ws, col_letter, start_row, end_row):
    """Apply green/amber/red/gray conditional formatting to status column."""
    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator="equal", formula=['"Compliant"'],
        fill=PatternFill(start_color=C_GREEN, end_color=C_GREEN, fill_type="solid")))
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator="equal", formula=['"Partial"'],
        fill=PatternFill(start_color=C_AMBER, end_color=C_AMBER, fill_type="solid")))
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator="equal", formula=['"Non-Compliant"'],
        fill=PatternFill(start_color=C_RED, end_color=C_RED, fill_type="solid")))
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator="equal", formula=['"N/A"'],
        fill=PatternFill(start_color=C_LGRAY, end_color=C_LGRAY, fill_type="solid")))


# =============================================================================
# PUBLIC API — called by each thin generator
# =============================================================================

def generate_checklist(document_id, control_id, control_name, source_policy,
                       requirements, output_filename=None,
                       iso_standard="ISO/IEC 27701:2025"):
    """Generate a complete Privacy/Cloud Compliance Checklist workbook.

    Args:
        document_id:    e.g. "PRIV-CHK-A.1.2.2-5"
        control_id:     e.g. "A.1.2.2-5"
        control_name:   e.g. "Lawful Basis and Consent"
        source_policy:  e.g. "PRIV-POL-A.1.2.2-5"
        requirements:   OrderedDict of domain_name -> [(req_id, section, text), ...]
        output_filename: override output path (default: auto-generated)
        iso_standard:   ISO standard string for the control_ref line on domain sheets
                        (default: "ISO/IEC 27701:2025"; use "ISO/IEC 27018:2025" for Cloud)

    Returns:
        0 on success, 1 on failure
    """
    if output_filename is None:
        ts = datetime.now().strftime("%Y%m%d")
        output_filename = f"{document_id}_Compliance_Checklist_{ts}.xlsx"

    control_ref = f"{iso_standard} \u2014 Controls {control_id}: {control_name}"

    logger.info("=" * 70)
    logger.info(f"  {document_id} \u2014 {control_name} Compliance Checklist Generator")
    logger.info(f"  Source: {source_policy}")
    logger.info("=" * 70)

    try:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create all sheets
        sheet_names = ["Executive Summary", "Dashboard"]
        sheet_names.extend(requirements.keys())
        for name in sheet_names:
            wb.create_sheet(title=name)

        styles = setup_styles()

        # Domain sheets first (Dashboard references them)
        domain_config = []
        for domain_name, reqs in requirements.items():
            ws = wb[domain_name]
            last_row = create_domain_sheet(ws, styles, domain_name, reqs, control_ref)
            domain_config.append((domain_name, 4, last_row))

        create_dashboard(wb["Dashboard"], styles, domain_config, control_name)
        create_executive_summary(wb["Executive Summary"], styles, domain_config,
                                 document_id, control_name, source_policy)

        total_reqs = sum(len(reqs) for reqs in requirements.values())
        wb.save(output_filename)

        logger.info("")
        logger.info("=" * 70)
        logger.info(f"  SUCCESS: {output_filename}")
        logger.info(f"  Sheets: {len(wb.sheetnames)} (Exec Summary + Dashboard + {len(requirements)} domains)")
        logger.info(f"  Requirements: {total_reqs} 'shall' statements from {source_policy}")
        logger.info("=" * 70)
        return 0

    except Exception as e:
        logger.error(f"Generation failed: {e}")
        return 1
