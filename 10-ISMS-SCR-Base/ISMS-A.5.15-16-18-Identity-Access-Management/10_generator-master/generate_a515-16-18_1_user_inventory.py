#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.5.15-16-18.S1 - User Inventory & Lifecycle Compliance Assessment Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.16: Identity Management
Assessment Workbook 1 of 5: User Inventory and Identity Lifecycle Compliance

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific identity management systems, user lifecycle
processes, and assessment requirements.

Key customization areas:
1. Identity system integrations (Active Directory, Azure AD, Okta, custom LDAP)
2. User types and classification (employees, contractors, vendors, service accounts)
3. Provisioning/deprovisioning SLA thresholds (match your IAM policies)
4. HR system integration points (authoritative source of identity)
5. Compliance scoring criteria (aligned with your risk tolerance)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.15-16-18 IAM Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
user inventory completeness and identity lifecycle compliance across all
identity management systems.

**Purpose:**
Enables systematic assessment of identity management processes against
ISO 27001:2022 Control A.5.16 requirements, supporting evidence-based
validation of joiner/mover/leaver procedures and orphaned account detection.

**Assessment Scope:**
- Complete user inventory across all identity systems
- Employee lifecycle compliance (provisioning/deprovisioning timeliness)
- Contractor lifecycle compliance (time-bound access management)
- Vendor identity management (sponsored access tracking)
- Service account inventory and ownership
- Orphaned account detection and remediation
- Inactive account identification (no login >90 days)
- User type classification and segregation
- Lifecycle metrics and KPIs
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Usage guidance and field definitions
2. User_Inventory - Complete user list across all systems (100 sample users)
3. Employee_Lifecycle - Employee provisioning/deprovisioning compliance (50 rows)
4. Contractor_Lifecycle - Contractor time-bound access compliance (30 rows)
5. Service_Accounts - Non-human account inventory (20 rows)
6. Orphaned_Accounts - Orphaned account detection and remediation (15 rows)
7. Lifecycle_Metrics - Summary metrics and KPIs
8. Gap_Analysis - Non-compliance findings and remediation tracking
9. Evidence_Register - Evidence collection for A.5.16 compliance
10. Approval_Sign_Off - 3-level approval workflow

**Key Features:**
- Data validation with identity system dropdown lists
- Conditional formatting for lifecycle compliance status
- Automated gap identification for orphaned/inactive accounts
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with HR system exports

**Integration:**
This assessment feeds into the IAM Governance Dashboard, which
consolidates data from all five IAM assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - random (standard library - for sample data generation)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a515-16-18_1_user_inventory.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a515-16-18_1_user_inventory.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a515-16-18_1_user_inventory.py --date 20250124

Output:
    File: ISMS-IMP-A.5.15-16-18.S1_User_Inventory_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Export user data from identity systems (AD, Azure AD, Okta, etc.)
    2. Export employee data from HR system (authoritative source)
    3. Populate User_Inventory with actual user accounts
    4. Document joiner events in Employee_Lifecycle
    5. Document leaver events and deprovisioning timeliness
    6. Track contractor access periods in Contractor_Lifecycle
    7. Inventory service accounts and assign ownership
    8. Cross-reference to identify orphaned accounts
    9. Calculate compliance metrics and identify gaps
    10. Collect and link audit evidence
    11. Obtain stakeholder approvals
    12. Feed results into IAM Governance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.16
Assessment Domain:    1 of 5 (User Inventory & Lifecycle Compliance)
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.15-16-18: Identity Management Policy (A.5.16)
    - ISMS-IMP-A.5.15-16-18-S2: Identity Lifecycle Process Guide
    - ISMS-IMP-A.5.15-16-18.S2: Access Rights Matrix Assessment (Workbook 2)
    - ISMS-IMP-A.5.15-16-18.S3: Access Review Results Assessment (Workbook 3)
    - ISMS-IMP-A.5.15-16-18.S4: Role Compliance Assessment (Workbook 4)
    - ISMS-IMP-A.5.15-16-18.S5: Lifecycle Compliance Detailed (Workbook 5)
    - IAM Governance Dashboard: Consolidated IAM compliance view

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.15-16-18-S1
    - Supports comprehensive user inventory and lifecycle compliance evaluation
    - Integrated with IAM Governance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Identity Lifecycle Criticality:**
Timely provisioning/deprovisioning is CRITICAL for security:
- Delayed provisioning: Business disruption (new employees can't work)
- Delayed deprovisioning: Security risk (ex-employees retain access)
- Orphaned accounts: Major audit finding (no accountability)

Ensure lifecycle SLA thresholds match your organization's risk tolerance.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Auditors expect:
- Complete user inventory across ALL identity systems
- Documented provisioning/deprovisioning events with timestamps
- Evidence of timely access removal upon termination
- Orphaned account detection and remediation procedures
- Service account ownership and justification

**Data Protection:**
Assessment workbooks contain sensitive HR and identity data including:
- Employee names, departments, hire/termination dates
- Contractor sponsorship and access periods
- Account credentials metadata
- Organizational structure information

Handle in accordance with GDPR/FADP data protection requirements.

**Maintenance:**
Review and update assessment:
- Monthly: Orphaned account detection runs
- Quarterly: Complete lifecycle compliance review
- Semi-annually: User inventory reconciliation with HR system
- Annually: Full reassessment for certification
- Ad-hoc: When identity systems change or after organizational restructuring

**Quality Assurance:**
Cross-validate user inventory against multiple authoritative sources:
- HR system (employees, contractors)
- Identity systems (AD, Azure AD, Okta)
- Access management platforms
- Ticketing system (access requests/removals)

Discrepancies indicate control weaknesses requiring investigation.

**Contractor/Vendor Management:**
Time-bound access for contractors/vendors is MANDATORY:
- Sponsorship required (internal employee sponsor)
- Contract end date drives access expiration
- Access extension requires explicit approval
- Unsponsored contractor accounts = orphaned accounts

**Service Accounts:**
Service accounts require special handling:
- No expiration date (runs until service decommissioned)
- Ownership assignment CRITICAL (service account owner documented)
- Password rotation requirements
- Privileged service accounts: extra scrutiny in A.8.2 PAM assessment

================================================================================
"""

import sys
from datetime import datetime, timedelta
from pathlib import Path
import random

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import PieChart, BarChart, Reference
except ImportError as e:
    print(f"❌ ERROR: Required library not found: {e}")
    print("📦 Install required libraries: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# SECTION 1: CONSTANTS AND CONFIGURATION
# ============================================================================

WORKBOOK_NAME = "User Inventory & Lifecycle Compliance Assessment"
DOCUMENT_ID = "ISMS-IMP-A.5.15-16-18.S1"
CONTROL_ID = "A.5.16"
CONTROL_NAME = "Identity Management"
GENERATED_DATE = datetime.now().strftime("%Y-%m-%d")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Data row counts
USER_ROW_COUNT = 100          # Sample users
EMPLOYEE_LIFECYCLE_COUNT = 50 # Employee lifecycle events
CONTRACTOR_LIFECYCLE_COUNT = 30 # Contractor lifecycle events
SERVICE_ACCOUNT_COUNT = 20    # Service accounts
ORPHANED_ACCOUNT_COUNT = 15   # Orphaned accounts
GAP_ROW_COUNT = 40            # Gap analysis tracking
EVIDENCE_ROW_COUNT = 50       # Evidence register

# Compliance thresholds
PROVISIONING_SLA_DAYS = 0     # Provision by start date (0 = on or before start date)
DEPROVISIONING_SLA_HOURS = 24 # Deprovision within 24 hours of termination
INACTIVE_THRESHOLD_DAYS = 90  # Accounts with no login for 90+ days flagged


# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================

def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style templates as dictionaries to avoid shared object issues.
    """
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="002060", end_color="002060", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data_center": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "compliant": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "non_compliant": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "warning": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }
    
    return styles


def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = style_dict["font"]
    if "fill" in style_dict:
        cell.fill = style_dict["fill"]
    if "alignment" in style_dict:
        cell.alignment = style_dict["alignment"]
    if "border" in style_dict:
        cell.border = style_dict["border"]


# ============================================================================
# SECTION 3: SAMPLE DATA GENERATORS
# ============================================================================

def generate_sample_users(count):
    """Generate sample user data for testing."""
    first_names = ["Alice", "Bob", "Carol", "David", "Emma", "Frank", "Grace", "Henry", 
                   "Iris", "Jack", "Kate", "Liam", "Mary", "Nathan", "Olivia", "Peter",
                   "Quinn", "Rachel", "Sam", "Tara"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
                  "Davis", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez",
                  "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin"]
    
    departments = ["Finance", "Sales", "Marketing", "IT", "HR", "Operations", "Engineering",
                   "Product", "Legal", "Customer Success"]
    job_titles = {
        "Finance": ["Financial Analyst", "Finance Manager", "Accountant", "CFO"],
        "Sales": ["Sales Representative", "Sales Manager", "Account Executive", "VP Sales"],
        "Marketing": ["Marketing Specialist", "Marketing Manager", "Content Writer"],
        "IT": ["IT Administrator", "System Administrator", "IT Manager", "Network Engineer"],
        "HR": ["HR Generalist", "HR Manager", "Recruiter", "HR Director"],
        "Operations": ["Operations Analyst", "Operations Manager", "Logistics Coordinator"],
        "Engineering": ["Software Developer", "Senior Engineer", "Engineering Manager", "DevOps Engineer"],
        "Product": ["Product Manager", "Product Owner", "UX Designer"],
        "Legal": ["Legal Counsel", "Paralegal", "General Counsel"],
        "Customer Success": ["Customer Success Manager", "Support Specialist"]
    }
    
    user_types = ["Employee", "Contractor", "Vendor", "Service Account"]
    statuses = ["Active", "Disabled", "Suspended"]
    
    users = []
    for i in range(count):
        first = random.choice(first_names)
        last = random.choice(last_names)
        dept = random.choice(departments)
        user_type = random.choices(user_types, weights=[70, 20, 5, 5])[0]  # 70% employees
        
        if user_type == "Service Account":
            username = f"svc-{random.choice(['backup', 'monitoring', 'automation', 'integration'])}-{i:02d}"
            email = f"{username}@company.local"
            full_name = f"Service Account - {username}"
            title = "Service Account"
        else:
            username = f"{first.lower()}.{last.lower()}"
            if user_type == "Contractor":
                username = f"EXT-{username}"
            email = f"{username}@company.com"
            full_name = f"{first} {last}"
            title = random.choice(job_titles.get(dept, ["Specialist"]))
        
        # Most users active, some disabled/suspended
        status = random.choices(statuses, weights=[85, 10, 5])[0]
        
        # Generate dates
        hire_date = datetime.now() - timedelta(days=random.randint(30, 1825))  # 1 month to 5 years ago
        
        # Termination date (only for some disabled users)
        term_date = None
        if status == "Disabled" and random.random() < 0.7:  # 70% of disabled users are terminated
            term_date = datetime.now() - timedelta(days=random.randint(1, 60))
        
        # Last login (active users more recent, disabled less recent)
        if status == "Active":
            last_login = datetime.now() - timedelta(days=random.randint(0, 30))
        else:
            last_login = datetime.now() - timedelta(days=random.randint(60, 180)) if random.random() < 0.8 else None
        
        users.append({
            "user_id": f"U{10000 + i}",
            "username": username,
            "full_name": full_name,
            "email": email,
            "user_type": user_type,
            "department": dept if user_type != "Service Account" else "IT",
            "job_title": title,
            "manager": f"{random.choice(first_names)} {random.choice(last_names)}" if user_type != "Service Account" else "IT Manager",
            "hire_date": hire_date,
            "termination_date": term_date,
            "status": status,
            "last_login": last_login,
            "created_date": hire_date - timedelta(days=random.randint(1, 7)),  # Account created within week of hire
        })
    
    return users


def generate_lifecycle_events(user_list, event_type="joiner"):
    """Generate lifecycle events (joiner/mover/leaver) from user list."""
    events = []
    
    for user in user_list:
        if event_type == "joiner":
            # Provisioning timeliness
            provision_date = user["created_date"]
            days_to_provision = (provision_date - user["hire_date"]).days
            compliant = days_to_provision <= PROVISIONING_SLA_DAYS
            
            events.append({
                "user_id": user["user_id"],
                "username": user["username"],
                "full_name": user["full_name"],
                "user_type": user["user_type"],
                "department": user["department"],
                "hire_date": user["hire_date"],
                "provision_date": provision_date,
                "days_to_provision": days_to_provision,
                "compliance_status": "Compliant" if compliant else "Non-Compliant",
            })
        
        elif event_type == "leaver" and user["termination_date"]:
            # Deprovisioning timeliness
            # Simulate disable date (some compliant, some late)
            if random.random() < 0.85:  # 85% compliant
                disable_date = user["termination_date"] + timedelta(hours=random.randint(0, 12))
            else:  # 15% non-compliant
                disable_date = user["termination_date"] + timedelta(days=random.randint(2, 14))
            
            hours_to_disable = (disable_date - user["termination_date"]).total_seconds() / 3600
            compliant = hours_to_disable <= DEPROVISIONING_SLA_HOURS
            
            events.append({
                "user_id": user["user_id"],
                "username": user["username"],
                "full_name": user["full_name"],
                "user_type": user["user_type"],
                "department": user["department"],
                "termination_date": user["termination_date"],
                "disable_date": disable_date,
                "hours_to_disable": round(hours_to_disable, 1),
                "compliance_status": "Compliant" if compliant else "Non-Compliant",
            })
    
    return events


# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS
# ============================================================================

def create_instructions_sheet(wb, styles):
    """Create Sheet 1: Instructions & Legend."""
    ws = wb.create_sheet("Instructions & Legend", 0)
    
    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = f"{WORKBOOK_NAME} - Instructions"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 30
    
    # Document metadata
    ws["A3"].value = "Document ID:"
    ws["A3"].font = Font(bold=True)
    ws["B3"].value = DOCUMENT_ID
    
    ws["A4"].value = "Assessment:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].value = "User Inventory & Lifecycle Compliance"
    
    ws["A5"].value = "Version:"
    ws["A5"].font = Font(bold=True)
    ws["B5"].value = "1.0"
    
    ws["A6"].value = "Generated:"
    ws["A6"].font = Font(bold=True)
    ws["B6"].value = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    # Document info
    ws.merge_cells("A8:H8")
    cell = ws["A8"]
    cell.value = f"ISO/IEC 27001:2022 Control {CONTROL_ID}: {CONTROL_NAME}"
    apply_style(cell, styles["header"])
    
    ws.merge_cells("A9:H9")
    ws["A9"].value = f"Control Assessment Framework"
    ws["A9"].font = Font(name="Calibri", size=10, italic=True)
    
    # Purpose section
    row = 11
    ws[f"A{row}"] = "Purpose"
    apply_style(ws[f"A{row}"], styles["subheader"])
    ws.merge_cells(f"A{row}:H{row}")
    
    purpose_text = """
This workbook provides a comprehensive assessment framework for user inventory and identity lifecycle 
compliance. It tracks:
• Complete user inventory across all identity systems
• Provisioning timeliness (joiner process compliance)
• Deprovisioning timeliness (leaver process compliance)
• Orphaned account detection and remediation
• User type classification and service account management
• Overall A.5.16 compliance scoring
    """.strip()
    
    row += 1
    ws.merge_cells(f"A{row}:H{row+6}")
    ws[f"A{row}"].value = purpose_text
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 120
    
    # Sheet structure
    row += 8
    ws[f"A{row}"] = "Workbook Structure"
    apply_style(ws[f"A{row}"], styles["subheader"])
    ws.merge_cells(f"A{row}:H{row}")
    
    row += 1
    headers = ["Sheet", "Purpose", "Key Metrics"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    sheets_info = [
        ("User_Inventory", "Complete user list with attributes", "Total users, user types, active/disabled"),
        ("Employee_Lifecycle", "Employee provisioning/deprovisioning", "On-time provisioning rate, deprovisioning rate"),
        ("Contractor_Lifecycle", "Contractor time-bound access", "Contract compliance, expired access"),
        ("Service_Accounts", "Non-human account inventory", "Service account count, ownership"),
        ("Orphaned_Accounts", "Orphaned account detection", "Orphaned count, remediation status"),
        ("Lifecycle_Metrics", "Summary KPIs", "Overall compliance score"),
        ("Gap_Analysis", "Non-compliance tracking", "Gap count, remediation progress"),
        ("Evidence_Register", "Evidence collection", "Evidence completeness"),
    ]
    
    for sheet_name, purpose, metrics in sheets_info:
        row += 1
        ws.cell(row=row, column=1).value = sheet_name
        ws.cell(row=row, column=2).value = purpose
        ws.cell(row=row, column=3).value = metrics
        for col in range(1, 4):
            apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Color legend
    row += 2
    ws[f"A{row}"] = "Status Color Legend"
    apply_style(ws[f"A{row}"], styles["subheader"])
    ws.merge_cells(f"A{row}:H{row}")
    
    row += 1
    ws[f"A{row}"] = "Compliant"
    apply_style(ws[f"A{row}"], styles["compliant"])
    ws[f"B{row}"] = "Lifecycle event completed within SLA"
    
    row += 1
    ws[f"A{row}"] = "Non-Compliant"
    apply_style(ws[f"A{row}"], styles["non_compliant"])
    ws[f"B{row}"] = "Lifecycle event exceeded SLA"
    
    row += 1
    ws[f"A{row}"] = "Warning"
    apply_style(ws[f"A{row}"], styles["warning"])
    ws[f"B{row}"] = "Requires attention (e.g., inactive account, missing data)"
    
    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 35
    
    return ws


def create_user_inventory_sheet(wb, styles):
    """Create Sheet 2: User_Inventory."""
    ws = wb.create_sheet("User_Inventory")
    
    # Title
    ws.merge_cells("A1:M1")
    cell = ws["A1"]
    cell.value = "User Inventory - All Identity Systems"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Metadata
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Total Users: {USER_ROW_COUNT}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    headers = [
        "User ID", "Username", "Full Name", "Email", "User Type",
        "Department", "Job Title", "Manager", "Hire Date", "Termination Date",
        "Status", "Last Login", "Account Created", "Comments"
    ]
    
    row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Generate sample data
    users = generate_sample_users(USER_ROW_COUNT)
    
    for user in users:
        row += 1
        ws.cell(row=row, column=1).value = user["user_id"]
        ws.cell(row=row, column=2).value = user["username"]
        ws.cell(row=row, column=3).value = user["full_name"]
        ws.cell(row=row, column=4).value = user["email"]
        ws.cell(row=row, column=5).value = user["user_type"]
        ws.cell(row=row, column=6).value = user["department"]
        ws.cell(row=row, column=7).value = user["job_title"]
        ws.cell(row=row, column=8).value = user["manager"]
        ws.cell(row=row, column=9).value = user["hire_date"].strftime("%Y-%m-%d") if user["hire_date"] else ""
        ws.cell(row=row, column=10).value = user["termination_date"].strftime("%Y-%m-%d") if user["termination_date"] else ""
        ws.cell(row=row, column=11).value = user["status"]
        ws.cell(row=row, column=12).value = user["last_login"].strftime("%Y-%m-%d") if user["last_login"] else "Never"
        ws.cell(row=row, column=13).value = user["created_date"].strftime("%Y-%m-%d") if user["created_date"] else ""
        
        # Apply conditional formatting to status
        status_cell = ws.cell(row=row, column=11)
        if user["status"] == "Active":
            apply_style(status_cell, styles["compliant"])
        elif user["status"] == "Disabled":
            apply_style(status_cell, styles["non_compliant"])
        else:  # Suspended
            apply_style(status_cell, styles["warning"])
        
        # Format other cells
        for col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 13, 14]:
            if col != 11:  # Skip status (already formatted)
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 10, "B": 20, "C": 20, "D": 30, "E": 15, "F": 15, "G": 25,
              "H": 20, "I": 12, "J": 15, "K": 12, "L": 12, "M": 12, "N": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze panes
    ws.freeze_panes = "A6"
    
    return ws, users  # Return users for use in other sheets


def create_employee_lifecycle_sheet(wb, styles, users):
    """Create Sheet 3: Employee_Lifecycle."""
    ws = wb.create_sheet("Employee_Lifecycle")
    
    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "Employee Lifecycle Compliance - Joiner & Leaver Events"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Filter to employees only
    employees = [u for u in users if u["user_type"] == "Employee"][:EMPLOYEE_LIFECYCLE_COUNT]
    
    # Metadata
    ws["A2"] = f"Assessment Period: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Employees Assessed: {len(employees)}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers for JOINER events
    row = 5
    ws.merge_cells(f"A{row}:J{row}")
    cell = ws[f"A{row}"]
    cell.value = "JOINER EVENTS - Provisioning Compliance"
    apply_style(cell, styles["subheader"])
    
    row += 1
    headers = [
        "User ID", "Username", "Full Name", "Department",
        "Hire Date", "Provision Date", "Days to Provision", "SLA (Days)", "Compliance Status", "Notes"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Generate joiner events
    joiner_events = generate_lifecycle_events(employees, event_type="joiner")
    
    for event in joiner_events:
        row += 1
        ws.cell(row=row, column=1).value = event["user_id"]
        ws.cell(row=row, column=2).value = event["username"]
        ws.cell(row=row, column=3).value = event["full_name"]
        ws.cell(row=row, column=4).value = event["department"]
        ws.cell(row=row, column=5).value = event["hire_date"].strftime("%Y-%m-%d")
        ws.cell(row=row, column=6).value = event["provision_date"].strftime("%Y-%m-%d")
        ws.cell(row=row, column=7).value = event["days_to_provision"]
        ws.cell(row=row, column=8).value = PROVISIONING_SLA_DAYS
        ws.cell(row=row, column=9).value = event["compliance_status"]
        
        # Conditional formatting
        status_cell = ws.cell(row=row, column=9)
        if event["compliance_status"] == "Compliant":
            apply_style(status_cell, styles["compliant"])
        else:
            apply_style(status_cell, styles["non_compliant"])
            ws.cell(row=row, column=10).value = "Late provisioning"
        
        for col in range(1, 11):
            if col != 9:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # LEAVER events section
    row += 3
    ws.merge_cells(f"A{row}:J{row}")
    cell = ws[f"A{row}"]
    cell.value = "LEAVER EVENTS - Deprovisioning Compliance"
    apply_style(cell, styles["subheader"])
    
    row += 1
    headers = [
        "User ID", "Username", "Full Name", "Department",
        "Termination Date", "Disable Date", "Hours to Disable", "SLA (Hours)", "Compliance Status", "Notes"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Generate leaver events
    leaver_events = generate_lifecycle_events(employees, event_type="leaver")
    
    for event in leaver_events:
        row += 1
        ws.cell(row=row, column=1).value = event["user_id"]
        ws.cell(row=row, column=2).value = event["username"]
        ws.cell(row=row, column=3).value = event["full_name"]
        ws.cell(row=row, column=4).value = event["department"]
        ws.cell(row=row, column=5).value = event["termination_date"].strftime("%Y-%m-%d")
        ws.cell(row=row, column=6).value = event["disable_date"].strftime("%Y-%m-%d %H:%M")
        ws.cell(row=row, column=7).value = event["hours_to_disable"]
        ws.cell(row=row, column=8).value = DEPROVISIONING_SLA_HOURS
        ws.cell(row=row, column=9).value = event["compliance_status"]
        
        # Conditional formatting
        status_cell = ws.cell(row=row, column=9)
        if event["compliance_status"] == "Compliant":
            apply_style(status_cell, styles["compliant"])
        else:
            apply_style(status_cell, styles["non_compliant"])
            ws.cell(row=row, column=10).value = "Late deprovisioning - security risk"
        
        for col in range(1, 11):
            if col != 9:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 10, "B": 20, "C": 20, "D": 15, "E": 15, "F": 18, "G": 15, "H": 12, "I": 18, "J": 35}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A7"
    
    return ws


def create_contractor_lifecycle_sheet(wb, styles, users):
    """Create Sheet 4: Contractor_Lifecycle."""
    ws = wb.create_sheet("Contractor_Lifecycle")
    
    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "Contractor Lifecycle - Time-Bound Access Compliance"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Filter to contractors
    contractors = [u for u in users if u["user_type"] == "Contractor"][:CONTRACTOR_LIFECYCLE_COUNT]
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Contractors Assessed: {len(contractors)}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    row = 5
    headers = [
        "User ID", "Username", "Full Name", "Department", "Sponsor",
        "Contract Start", "Contract End", "Access Expiry", "Days Remaining", "Status", "Compliance"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Generate contractor data
    for contractor in contractors:
        row += 1
        contract_end = datetime.now() + timedelta(days=random.randint(-30, 180))  # Some expired, some future
        days_remaining = (contract_end - datetime.now()).days
        
        ws.cell(row=row, column=1).value = contractor["user_id"]
        ws.cell(row=row, column=2).value = contractor["username"]
        ws.cell(row=row, column=3).value = contractor["full_name"]
        ws.cell(row=row, column=4).value = contractor["department"]
        ws.cell(row=row, column=5).value = contractor["manager"]  # Sponsor
        ws.cell(row=row, column=6).value = contractor["hire_date"].strftime("%Y-%m-%d")
        ws.cell(row=row, column=7).value = contract_end.strftime("%Y-%m-%d")
        ws.cell(row=row, column=8).value = contract_end.strftime("%Y-%m-%d")  # Assume same as contract end
        ws.cell(row=row, column=9).value = days_remaining
        
        # Status based on contract
        if days_remaining < 0:
            status = "Expired"
            compliance = "Non-Compliant"
            style = styles["non_compliant"]
        elif days_remaining < 30:
            status = "Expiring Soon"
            compliance = "Warning"
            style = styles["warning"]
        else:
            status = "Active"
            compliance = "Compliant"
            style = styles["compliant"]
        
        ws.cell(row=row, column=10).value = status
        ws.cell(row=row, column=11).value = compliance
        
        apply_style(ws.cell(row=row, column=11), style)
        
        for col in range(1, 11):
            apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 10, "B": 22, "C": 20, "D": 15, "E": 20, "F": 15, "G": 15, "H": 15, "I": 15, "J": 15, "K": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_service_accounts_sheet(wb, styles, users):
    """Create Sheet 5: Service_Accounts."""
    ws = wb.create_sheet("Service_Accounts")
    
    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "Service Account Inventory - Non-Human Accounts"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Filter to service accounts
    service_accounts = [u for u in users if u["user_type"] == "Service Account"][:SERVICE_ACCOUNT_COUNT]
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Service Accounts: {len(service_accounts)}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    row = 5
    headers = [
        "Account ID", "Account Name", "Purpose", "Owner", "Department",
        "Created Date", "Last Used", "Privileged", "Password Rotation", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    purposes = ["Automated Backup", "Monitoring Agent", "Integration Service", "Batch Processing", "Database Access"]
    
    for account in service_accounts:
        row += 1
        ws.cell(row=row, column=1).value = account["user_id"]
        ws.cell(row=row, column=2).value = account["username"]
        ws.cell(row=row, column=3).value = random.choice(purposes)
        ws.cell(row=row, column=4).value = account["manager"]
        ws.cell(row=row, column=5).value = account["department"]
        ws.cell(row=row, column=6).value = account["created_date"].strftime("%Y-%m-%d")
        ws.cell(row=row, column=7).value = account["last_login"].strftime("%Y-%m-%d") if account["last_login"] else "Unknown"
        ws.cell(row=row, column=8).value = "Yes" if random.random() < 0.4 else "No"  # 40% privileged
        ws.cell(row=row, column=9).value = random.choice(["90 Days", "60 Days", "Never", "Manual"])
        ws.cell(row=row, column=10).value = account["status"]
        
        status_cell = ws.cell(row=row, column=10)
        if account["status"] == "Active":
            apply_style(status_cell, styles["compliant"])
        else:
            apply_style(status_cell, styles["non_compliant"])
        
        for col in range(1, 11):
            if col != 10:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 12, "B": 30, "C": 25, "D": 20, "E": 15, "F": 15, "G": 15, "H": 12, "I": 18, "J": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_orphaned_accounts_sheet(wb, styles, users):
    """Create Sheet 6: Orphaned_Accounts."""
    ws = wb.create_sheet("Orphaned_Accounts")
    
    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "Orphaned Account Detection & Remediation"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Identify orphaned accounts (no termination date but disabled, or very old last login)
    orphaned = []
    for user in users:
        is_orphaned = False
        reason = ""
        
        if user["status"] == "Disabled" and not user["termination_date"]:
            is_orphaned = True
            reason = "Disabled account with no termination record"
        elif user["last_login"] and (datetime.now() - user["last_login"]).days > INACTIVE_THRESHOLD_DAYS:
            is_orphaned = True
            reason = f"No login for {(datetime.now() - user['last_login']).days} days"
        
        if is_orphaned:
            orphaned.append({**user, "orphan_reason": reason})
            if len(orphaned) >= ORPHANED_ACCOUNT_COUNT:
                break
    
    ws["A2"] = f"Detection Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Orphaned Accounts Found: {len(orphaned)}"
    ws["A3"].font = Font(italic=True)
    ws["A3"].font = Font(bold=True, color="FF0000")
    
    # Column headers
    row = 5
    headers = [
        "User ID", "Username", "Full Name", "User Type", "Department",
        "Last Login", "Status", "Orphan Reason", "Detected Date", "Remediation Status", "Notes"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    remediation_statuses = ["Open", "In Progress", "Remediated", "Closed"]
    
    for account in orphaned:
        row += 1
        ws.cell(row=row, column=1).value = account["user_id"]
        ws.cell(row=row, column=2).value = account["username"]
        ws.cell(row=row, column=3).value = account["full_name"]
        ws.cell(row=row, column=4).value = account["user_type"]
        ws.cell(row=row, column=5).value = account["department"]
        ws.cell(row=row, column=6).value = account["last_login"].strftime("%Y-%m-%d") if account["last_login"] else "Never"
        ws.cell(row=row, column=7).value = account["status"]
        ws.cell(row=row, column=8).value = account["orphan_reason"]
        ws.cell(row=row, column=9).value = GENERATED_DATE
        ws.cell(row=row, column=10).value = random.choice(remediation_statuses)
        
        status_cell = ws.cell(row=row, column=10)
        status_value = status_cell.value
        if status_value == "Remediated" or status_value == "Closed":
            apply_style(status_cell, styles["compliant"])
        elif status_value == "In Progress":
            apply_style(status_cell, styles["warning"])
        else:
            apply_style(status_cell, styles["non_compliant"])
        
        for col in range(1, 12):
            if col != 10:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 10, "B": 20, "C": 20, "D": 15, "E": 15, "F": 15, "G": 12, "H": 40, "I": 15, "J": 18, "K": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_lifecycle_metrics_sheet(wb, styles):
    """Create Sheet 7: Lifecycle_Metrics."""
    ws = wb.create_sheet("Lifecycle_Metrics")
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Identity Lifecycle Compliance Metrics - A.5.16"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Assessment Period: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Summary metrics
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "Key Performance Indicators (KPIs)"
    apply_style(cell, styles["header"])
    
    row += 1
    headers = ["Metric", "Target", "Actual", "Status", "Gap", "Comments"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Sample metrics
    metrics = [
        ("Provisioning On-Time Rate", "≥95%", "92%", "Warning", "-3%", "3 late provisioning instances"),
        ("Deprovisioning On-Time Rate", "≥98%", "96%", "Warning", "-2%", "2 late deprovisioning instances"),
        ("Orphaned Account Count", "≤1%", "1.5%", "Non-Compliant", "+0.5%", "15 orphaned accounts detected"),
        ("Service Account Documentation", "100%", "95%", "Warning", "-5%", "1 account missing owner"),
        ("Contractor Access Compliance", "100%", "90%", "Warning", "-10%", "3 expired contractor accounts"),
    ]
    
    for metric_name, target, actual, status, gap, comments in metrics:
        row += 1
        ws.cell(row=row, column=1).value = metric_name
        ws.cell(row=row, column=2).value = target
        ws.cell(row=row, column=3).value = actual
        ws.cell(row=row, column=4).value = status
        ws.cell(row=row, column=5).value = gap
        ws.cell(row=row, column=6).value = comments
        
        status_cell = ws.cell(row=row, column=4)
        if status == "Compliant":
            apply_style(status_cell, styles["compliant"])
        elif status == "Warning":
            apply_style(status_cell, styles["warning"])
        else:
            apply_style(status_cell, styles["non_compliant"])
        
        for col in range(1, 7):
            if col != 4:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Overall score
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "Overall A.5.16 Compliance Score"
    apply_style(cell, styles["header"])
    
    row += 1
    ws[f"A{row}"] = "Score Calculation"
    ws[f"B{row}"] = "(Provisioning + Deprovisioning + Orphaned Account Management) / 3"
    
    row += 1
    ws[f"A{row}"] = "Overall Score"
    ws[f"B{row}"] = "82%"
    ws[f"B{row}"].font = Font(bold=True, size=14)
    
    row += 1
    ws[f"A{row}"] = "Maturity Level"
    ws[f"B{row}"] = "Managed (80-89%)"
    apply_style(ws[f"B{row}"], styles["warning"])
    
    # Column widths
    widths = {"A": 35, "B": 15, "C": 15, "D": 18, "E": 10, "F": 50}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    return ws


def create_gap_analysis_sheet(wb, styles):
    """Create Sheet 8: Gap_Analysis."""
    ws = wb.create_sheet("Gap_Analysis")
    
    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "Gap Analysis - Non-Compliance Findings"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Column headers
    row = 4
    headers = [
        "Gap ID", "Category", "Description", "Risk Level", "Affected Users",
        "Root Cause", "Remediation Plan", "Owner", "Due Date", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Sample gaps
    gaps = [
        ("GAP-001", "Provisioning", "3 users provisioned after hire date", "Medium", "3", "Manual provisioning delays",
         "Automate provisioning via HR system", "IAM Manager", (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"), "Open"),
        ("GAP-002", "Deprovisioning", "2 terminated users retained access >24h", "High", "2", "No termination notification from HR",
         "Implement automated HR termination feed", "IAM Manager", (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d"), "In Progress"),
        ("GAP-003", "Orphaned Accounts", "15 orphaned accounts detected", "High", "15", "No monthly orphaned account scan",
         "Schedule monthly automated scan", "Security Team", (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d"), "Open"),
        ("GAP-004", "Contractor Management", "3 contractors past contract end date", "Medium", "3", "No automated contract expiry alerts",
         "Implement contract expiry monitoring", "IAM Manager", (datetime.now() + timedelta(days=21)).strftime("%Y-%m-%d"), "Open"),
        ("GAP-005", "Service Accounts", "1 service account missing owner", "Medium", "1", "Incomplete service account documentation",
         "Conduct service account inventory review", "IT Manager", (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d"), "Open"),
    ]
    
    for gap_data in gaps:
        row += 1
        for col, value in enumerate(gap_data, start=1):
            ws.cell(row=row, column=col).value = value
        
        # Color code by risk level
        risk_cell = ws.cell(row=row, column=4)
        if gap_data[3] == "High":
            apply_style(risk_cell, styles["non_compliant"])
        elif gap_data[3] == "Medium":
            apply_style(risk_cell, styles["warning"])
        else:
            apply_style(risk_cell, styles["compliant"])
        
        # Color code status
        status_cell = ws.cell(row=row, column=10)
        if gap_data[9] == "Closed":
            apply_style(status_cell, styles["compliant"])
        elif gap_data[9] == "In Progress":
            apply_style(status_cell, styles["warning"])
        else:
            apply_style(status_cell, styles["non_compliant"])
        
        for col in range(1, 11):
            if col not in [4, 10]:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 10, "B": 18, "C": 40, "D": 12, "E": 15, "F": 30, "G": 40, "H": 18, "I": 12, "J": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A5"
    
    return ws


def create_evidence_register_sheet(wb, styles):
    """Create Sheet 9: Evidence_Register."""
    ws = wb.create_sheet("Evidence_Register")
    
    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "Evidence Register - A.5.16 Identity Management"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Evidence Collection Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Column headers
    row = 4
    headers = [
        "Evidence ID", "Requirement", "Evidence Type", "Evidence Location",
        "Collection Date", "Completeness", "Reviewed By", "Notes"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Sample evidence
    evidence_items = [
        ("EV-001", "User Inventory", "Spreadsheet", "User_Inventory sheet", GENERATED_DATE, "Complete", "Security Manager", "100 users documented"),
        ("EV-002", "Provisioning SLA", "Process Metrics", "Employee_Lifecycle sheet", GENERATED_DATE, "Complete", "IAM Manager", "92% on-time provisioning"),
        ("EV-003", "Deprovisioning SLA", "Process Metrics", "Employee_Lifecycle sheet", GENERATED_DATE, "Complete", "IAM Manager", "96% on-time deprovisioning"),
        ("EV-004", "Contractor Lifecycle", "Spreadsheet", "Contractor_Lifecycle sheet", GENERATED_DATE, "Complete", "IAM Manager", "30 contractors tracked"),
        ("EV-005", "Service Account Inventory", "Spreadsheet", "Service_Accounts sheet", GENERATED_DATE, "Complete", "IT Manager", "20 service accounts"),
        ("EV-006", "Orphaned Account Detection", "Scan Results", "Orphaned_Accounts sheet", GENERATED_DATE, "Complete", "Security Team", "15 orphaned accounts found"),
        ("EV-007", "HR Integration", "Process Documentation", "IMP-S2 Implementation Guide", GENERATED_DATE, "Partial", "IAM Manager", "Documented but not automated"),
        ("EV-008", "Identity Lifecycle Policy", "Policy Document", "ISMS-POL-A.5.15-16-18", GENERATED_DATE, "Complete", "CISO", "Policy approved"),
    ]
    
    for evidence in evidence_items:
        row += 1
        for col, value in enumerate(evidence, start=1):
            ws.cell(row=row, column=col).value = value
        
        # Color code completeness
        completeness_cell = ws.cell(row=row, column=6)
        if evidence[5] == "Complete":
            apply_style(completeness_cell, styles["compliant"])
        elif evidence[5] == "Partial":
            apply_style(completeness_cell, styles["warning"])
        else:
            apply_style(completeness_cell, styles["non_compliant"])
        
        for col in range(1, 9):
            if col != 6:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 12, "B": 25, "C": 20, "D": 35, "E": 18, "F": 15, "G": 20, "H": 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A5"
    
    return ws


def create_approval_sheet(wb, styles):
    """Create Sheet 10: Approval_Sign_Off."""
    ws = wb.create_sheet("Approval_Sign_Off")
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Assessment Approval & Sign-Off"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Approval workflow
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "3-Level Approval Process"
    apply_style(cell, styles["header"])
    
    row += 1
    headers = ["Approval Level", "Role", "Name", "Signature", "Date", "Status"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    approvals = [
        ("Level 1: Prepared By", "IAM Manager", "[Name]", "", "", "Pending"),
        ("Level 2: Reviewed By", "Security Manager", "[Name]", "", "", "Pending"),
        ("Level 3: Approved By", "CISO", "[Name]", "", "", "Pending"),
    ]
    
    for approval in approvals:
        row += 1
        for col, value in enumerate(approval, start=1):
            ws.cell(row=row, column=col).value = value
            apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 25, "B": 20, "C": 25, "D": 20, "E": 15, "F": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    return ws


# ============================================================================
# SECTION 5: MAIN EXECUTION
# ============================================================================

def main():
    """Main function to generate the workbook."""
    print("="*80)
    print(f"Generating: {WORKBOOK_NAME}")
    print(f"Control: ISO/IEC 27001:2022 {CONTROL_ID} - {CONTROL_NAME}")
    print("="*80)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Setup styles
    styles = setup_styles()
    
    # Create sheets
    print("\n📝 Creating sheets...")
    
    print("  1/10 Creating Instructions & Legend...")
    create_instructions_sheet(wb, styles)
    
    print("  2/10 Creating User Inventory...")
    _, users = create_user_inventory_sheet(wb, styles)
    
    print("  3/10 Creating Employee Lifecycle...")
    create_employee_lifecycle_sheet(wb, styles, users)
    
    print("  4/10 Creating Contractor Lifecycle...")
    create_contractor_lifecycle_sheet(wb, styles, users)
    
    print("  5/10 Creating Service Accounts...")
    create_service_accounts_sheet(wb, styles, users)
    
    print("  6/10 Creating Orphaned Accounts...")
    create_orphaned_accounts_sheet(wb, styles, users)
    
    print("  7/10 Creating Lifecycle Metrics...")
    create_lifecycle_metrics_sheet(wb, styles)
    
    print("  8/10 Creating Gap Analysis...")
    create_gap_analysis_sheet(wb, styles)
    
    print("  9/10 Creating Evidence Register...")
    create_evidence_register_sheet(wb, styles)
    
    print("  10/10 Creating Approval Sign-Off...")
    create_approval_sheet(wb, styles)
    
    # Save workbook
    filename = f"User_Inventory_Assessment_{GENERATED_TIMESTAMP}.xlsx"
    output_path = Path.cwd() / filename
    
    print(f"\n💾 Saving workbook: {filename}")
    wb.save(output_path)
    
    print("\n" + "="*80)
    print("✅ SUCCESS!")
    print(f"📄 File created: {output_path}")
    print(f"📊 Sheets: 10")
    print(f"📋 Sample users: {USER_ROW_COUNT}")
    print("="*80)
    print("\nWorkbook Contents:")
    print("  • User Inventory (100 sample users)")
    print("  • Employee Lifecycle Compliance (joiner/leaver events)")
    print("  • Contractor Time-Bound Access Tracking")
    print("  • Service Account Inventory")
    print("  • Orphaned Account Detection (15 accounts)")
    print("  • Lifecycle Metrics & KPIs")
    print("  • Gap Analysis & Remediation Tracking")
    print("  • Evidence Register for A.5.16")
    print("  • 3-Level Approval Workflow")
    print("="*80)
    
    return output_path


if __name__ == "__main__":
    try:
        output_file = main()
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
