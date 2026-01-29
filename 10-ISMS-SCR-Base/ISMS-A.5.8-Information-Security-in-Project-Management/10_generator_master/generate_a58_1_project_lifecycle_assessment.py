#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ISMS-IMP-A.5.8.1 - Project Lifecycle Security Assessment Generator
ISO/IEC 27001:2022 Control A.5.8
Complete Professional Implementation
"""
import sys
from datetime import datetime, timedelta
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Error: pip3 install openpyxl"); sys.exit(1)

DOCUMENT_ID = "ISMS-IMP-A.5.8.1"
CONTROL_REF = "ISO/IEC 27001:2022 Control A.5.8"

def setup_styles():
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    return {
        "header": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)
        },
        "section_header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="305496", end_color="305496", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True)
        },
        "subsection_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="000000"),
            "fill": PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True)
        },
        "column_header": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border
        },
        "protected_cell": {
            "fill": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border
        },
        "border": border
    }

def create_workbook():
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    sheets = [
        "Instructions & Legend",
        "2. Project Classification", 
        "3. Initiation Phase",
        "4. Planning Phase",
        "5. Execution Phase",
        "6. Monitoring Phase",
        "7. Closure Phase",
        "8. Compliance Dashboard",
        "9. Evidence Register",
        "10. Sign-Off"
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb

def create_instructions_sheet(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{DOCUMENT_ID} - Project Lifecycle Security Assessment\\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40
    row = 3
    ws[f"A{row}"] = "DOCUMENT CONTROL"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    row += 2
    for label, value in [
        ("Document ID:", DOCUMENT_ID),
        ("Version:", "2.0"),
        ("Control Reference:", CONTROL_REF),
        ("Purpose:", "Assess security integration across project lifecycle (6 phases)")
    ]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 55

def create_classification_sheet(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "PHASE 0: PROJECT CLASSIFICATION\\nDetermine Risk Level (High/Medium/Low)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    row = 3
    ws[f"A{row}"] = "Project Name:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]
    row += 1
    ws[f"A{row}"] = "Project Manager:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]
    row += 2
    ws[f"A{row}"] = "CLASSIFICATION CRITERIA (Answer Yes/No):"
    ws[f"A{row}"].font = styles["subsection_header"]["font"]
    ws[f"A{row}"].fill = styles["subsection_header"]["fill"]
    row += 1
    criteria = [
        "Processes Restricted or Confidential data?",
        "External-facing system/service?",
        "Handles payment card data (PCI DSS)?",
        "Subject to regulatory requirements (GDPR, HIPAA, etc.)?",
        "Critical business function (RTO <4 hours)?",
        "Integrates with critical infrastructure?"
    ]
    validations = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(validations)
    for i, criterion in enumerate(criteria, 1):
        ws[f"A{row}"] = criterion
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        validations.add(ws[f"B{row}"])
        row += 1
    row += 1
    ws[f"A{row}"] = "CALCULATED RISK LEVEL:"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"B{row}"] = f'=IF(COUNTIF(B{row-7}:B{row-2},"Yes")>=4,"High",IF(COUNTIF(B{row-7}:B{row-2},"Yes")>=2,"Medium","Low"))'
    ws[f"B{row}"].font = Font(bold=True, size=14, color="FF0000")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 15

def create_phase_sheet(ws, styles, phase_name, activities):
    ws.merge_cells("A1:E1")
    ws["A1"] = phase_name.upper()
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    row = 3
    headers = ["Activity", "Status", "Completion Date", "Evidence Link", "Notes"]
    widths = [60, 15, 15, 40, 40]
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws[f"{col}{row}"].alignment = styles["column_header"]["alignment"]
        ws[f"{col}{row}"].border = styles["border"]
        ws.column_dimensions[col].width = width
    row += 1
    validations = DataValidation(
        type="list",
        formula1='"✅ Complete,🔄 In Progress,⚠️ Incomplete,❌ Not Done,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations)
    for activity in activities:
        ws[f"A{row}"] = activity
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws[f"A{row}"].border = styles["border"]
        for col in ['B', 'C', 'D', 'E']:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = styles["border"]
        validations.add(ws[f"B{row}"])
        ws[f"C{row}"].number_format = "DD.MM.YYYY"
        row += 1
    ws.freeze_panes = "A4"

def create_dashboard_sheet(ws, styles):
    ws.merge_cells("A1:D1")
    ws["A1"] = "COMPLIANCE DASHBOARD\\nAutomated Phase Completion Scores"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    row = 3
    ws[f"A{row}"] = "PROJECT INFORMATION"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    row += 2
    ws[f"A{row}"] = "Project Name:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "='2. Project Classification'!B3"
    row += 1
    ws[f"A{row}"] = "Risk Classification:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "='2. Project Classification'!B18"
    ws[f"B{row}"].font = Font(bold=True, size=12, color="FF0000")
    row += 1
    ws[f"A{row}"] = "Assessment Date:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "=TODAY()"
    ws[f"B{row}"].number_format = "DD.MM.YYYY"
    row += 2
    ws[f"A{row}"] = "PHASE COMPLETION SCORES"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    row += 1
    ws[f"A{row}"] = "Phase"
    ws[f"B{row}"] = "Completion %"
    for col in ['A', 'B']:
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
    row += 1
    phases = [
        ("Initiation", "'3. Initiation Phase'", 4, 9),
        ("Planning", "'4. Planning Phase'", 4, 11),
        ("Execution", "'5. Execution Phase'", 4, 12),
        ("Monitoring", "'6. Monitoring Phase'", 4, 8),
        ("Closure", "'7. Closure Phase'", 4, 9)
    ]
    for phase_name, sheet_ref, start_row, end_row in phases:
        ws[f"A{row}"] = phase_name
        ws[f"B{row}"] = f'=COUNTIF({sheet_ref}!B{start_row}:B{end_row},"✅ Complete")/(COUNTA({sheet_ref}!B{start_row}:B{end_row})-COUNTIF({sheet_ref}!B{start_row}:B{end_row},"N/A"))'
        ws[f"B{row}"].number_format = "0%"
        row += 1
    row += 1
    ws[f"A{row}"] = "OVERALL COMPLIANCE:"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"B{row}"] = f"=AVERAGE(B{row-5}:B{row-1})"
    ws[f"B{row}"].number_format = "0%"
    ws[f"B{row}"].font = Font(bold=True, size=14)
    ws[f"B{row}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15

def create_evidence_register(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "EVIDENCE REGISTER\\nAudit Evidence Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    row = 3
    headers = ["Evidence ID", "Phase", "Description", "Location/Path", "Date Collected", "Status"]
    widths = [12, 20, 50, 45, 15, 18]
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws[f"{col}{row}"].alignment = styles["column_header"]["alignment"]
        ws[f"{col}{row}"].border = styles["border"]
        ws.column_dimensions[col].width = width
    row += 1
    validations_phase = DataValidation(
        type="list",
        formula1='"Classification,Initiation,Planning,Execution,Monitoring,Closure"',
        allow_blank=False
    )
    ws.add_data_validation(validations_phase)
    validations_status = DataValidation(
        type="list",
        formula1='"Collected,Verified,Pending,Missing"',
        allow_blank=False
    )
    ws.add_data_validation(validations_status)
    for i in range(100):
        ws[f"A{row}"] = f"=TEXT(ROW()-3,\"EV-000\")"
        ws[f"A{row}"].font = Font(color="808080")
        ws[f"A{row}"].fill = styles["protected_cell"]["fill"]
        ws[f"A{row}"].border = styles["border"]
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = styles["border"]
        validations_phase.add(ws[f"B{row}"])
        validations_status.add(ws[f"F{row}"])
        ws[f"E{row}"].number_format = "DD.MM.YYYY"
        row += 1
    ws.freeze_panes = "A4"

def create_signoff_sheet(ws, styles):
    ws.merge_cells("A1:E1")
    ws["A1"] = "SIGN-OFF & APPROVAL\\nMulti-Stakeholder Approval Workflow"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    row = 3
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    row += 2
    ws[f"A{row}"] = "Project Name:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "='2. Project Classification'!B3"
    row += 1
    ws[f"A{row}"] = "Overall Compliance:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "='8. Compliance Dashboard'!B15"
    ws[f"B{row}"].number_format = "0%"
    ws[f"B{row}"].font = Font(bold=True)
    row += 2
    ws[f"A{row}"] = "APPROVALS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    row += 1
    headers = ["Role", "Name", "Signature", "Date", "Decision"]
    for col_idx, header in enumerate(headers, 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
    row += 1
    validations = DataValidation(
        type="list",
        formula1='"✅ Approved,⚠️ Conditional,❌ Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(validations)
    for role in ["Project Manager", "Security Team", "Final Approver"]:
        ws[f"A{row}"] = role
        ws[f"A{row}"].font = Font(bold=True)
        for col in ['B', 'C', 'D', 'E']:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = styles["border"]
        validations.add(ws[f"E{row}"])
        ws[f"D{row}"].number_format = "DD.MM.YYYY"
        row += 1
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20

def main():
    print("=" * 78)
    print(f"{DOCUMENT_ID} - Project Lifecycle Security Assessment Generator")
    print(CONTROL_REF)
    print("=" * 78)
    wb = create_workbook()
    styles = setup_styles()
    print("\\n[1/10] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    print("[2/10] Creating Project Classification...")
    create_classification_sheet(wb["2. Project Classification"], styles)
    print("[3/10] Creating Initiation Phase...")
    create_phase_sheet(wb["3. Initiation Phase"], styles, "Phase 1: Initiation", [
        "Identify security stakeholders and establish communication plan",
        "Conduct initial risk assessment and document key risks",
        "Define security requirements baseline",
        "Allocate security budget and resources",
        "Document security responsibilities in project charter",
        "Obtain Phase 1 gate approval from security team"
    ])
    print("[4/10] Creating Planning Phase...")
    create_phase_sheet(wb["4. Planning Phase"], styles, "Phase 2: Planning", [
        "Document detailed security requirements (link to A.5.8.2 Register)",
        "Conduct threat modeling and document threat scenarios",
        "Develop security test plan and define test cases",
        "Complete Data Protection Impact Assessment (DPIA) if required",
        "Conduct vendor security assessment if third-party components",
        "Define security monitoring and logging requirements",
        "Establish incident response procedures",
        "Obtain Phase 2 gate approval from security team"
    ])
    print("[5/10] Creating Execution Phase...")
    create_phase_sheet(wb["5. Execution Phase"], styles, "Phase 3: Execution", [
        "Execute security testing per test plan (SAST, DAST, etc.)",
        "Conduct penetration testing and document findings",
        "Complete vulnerability scans and remediate critical/high findings",
        "Review and approve security architecture/design",
        "Verify secure coding practices and code review completion",
        "Test security controls and verify implementation",
        "Document security configurations and hardening",
        "Update threat model with implementation changes",
        "Obtain Phase 3 gate approval from security team"
    ])
    print("[6/10] Creating Monitoring Phase...")
    create_phase_sheet(wb["6. Monitoring Phase"], styles, "Phase 4: Monitoring", [
        "Monitor ongoing compliance with security requirements",
        "Review and update risk assessments for changes",
        "Assess security impact of change requests",
        "Track security metrics and KPIs",
        "Obtain Phase 4 gate approval for major milestones"
    ])
    print("[7/10] Creating Closure Phase...")
    create_phase_sheet(wb["7. Closure Phase"], styles, "Phase 5: Closure", [
        "Complete security handover documentation",
        "Verify all security testing complete and passed",
        "Document residual risks and obtain risk acceptance",
        "Conduct lessons learned session with security team",
        "Archive security documentation and evidence",
        "Obtain final security sign-off for production deployment"
    ])
    print("[8/10] Creating Compliance Dashboard...")
    create_dashboard_sheet(wb["8. Compliance Dashboard"], styles)
    print("[9/10] Creating Evidence Register...")
    create_evidence_register(wb["9. Evidence Register"], styles)
    print("[10/10] Creating Sign-Off sheet...")
    create_signoff_sheet(wb["10. Sign-Off"], styles)
    filename = f"{DOCUMENT_ID}_Project_Lifecycle_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"\\n✅ SUCCESS: {filename}")
    print(f"📄 File: {filename}")
    print(f"📊 Sheets: 10 sheets created")
    print("\\nNext Steps:")
    print("1. Complete Project Classification to determine risk level")
    print("2. Complete phases sequentially as project progresses")
    print("3. Link evidence in Evidence Register")
    print("4. Review Compliance Dashboard for scores")
    print("5. Obtain sign-offs when complete")
    print("=" * 78)

if __name__ == "__main__":
    main()
