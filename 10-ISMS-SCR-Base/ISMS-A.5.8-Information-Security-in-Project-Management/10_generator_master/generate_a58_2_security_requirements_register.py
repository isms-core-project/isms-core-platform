#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ISMS-IMP-A.5.8.2 - Security Requirements Register Generator
ISO/IEC 27001:2022 Control A.5.8
Complete Professional Implementation - ALL 13 Sheets
"""
import sys
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Error: pip3 install openpyxl"); sys.exit(1)

DOCUMENT_ID = "ISMS-IMP-A.5.8.2"
CONTROL_REF = "ISO/IEC 27001:2022 Control A.5.8"

def setup_styles():
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    return {
        "header": {"font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
                  "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
                  "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)},
        "section_header": {"font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
                          "fill": PatternFill(start_color="305496", end_color="305496", fill_type="solid"),
                          "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True)},
        "column_header": {"font": Font(name="Calibri", size=11, bold=True),
                         "fill": PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"),
                         "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
                         "border": border},
        "input_cell": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                      "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
                      "border": border},
        "example_cell": {"font": Font(italic=True, color="808080"),
                        "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
                        "border": border},
        "border": border
    }

def create_workbook():
    wb = Workbook()
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
    for name in ["Instructions & Legend", "Requirements Register", "App Security Examples",
                 "Data Protection Examples", "Access Control Examples", "Infrastructure Examples",
                 "Third-Party Examples", "Compliance Examples", "Traceability Matrix",
                 "Verification Checklist", "Gap Analysis", "Evidence Register", "Sign-Off"]:
        wb.create_sheet(title=name)
    return wb

def create_instructions_sheet(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{DOCUMENT_ID} - Security Requirements Register\\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40
    row = 3
    ws[f"A{row}"] = "PURPOSE"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    row += 2
    ws.merge_cells(f"A{row}:F{row+4}")
    ws[f"A{row}"] = ("Document and track security requirements across 6 categories:\\n\\n"
                     "1. Application Security - Input validation, authentication, session mgmt\\n"
                     "2. Data Protection - Encryption, classification, retention, backup\\n"
                     "3. Access Control - Authentication, authorization, privileged access\\n"
                     "4. Infrastructure - Network segmentation, hardening, patching\\n"
                     "5. Third-Party - Vendor assessment, contracts, API security\\n"
                     "6. Compliance - GDPR, nDSG, PCI DSS, regulations\\n\\n"
                     "13-FIELD MODEL: ID, Category, Statement, Source, Priority, Acceptance Criteria, "
                     "Implementation Status, Assigned To, Target Date, Verification Method, Test Status, "
                     "Evidence Link, Notes")
    ws[f"A{row}"].alignment = Alignment(wrap_text=True)
    row += 6
    ws[f"A{row}"] = "INSTRUCTIONS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    row += 1
    for i, instr in enumerate([
        "1. Review example requirements in Category sheets (Sheets 3-8)",
        "2. Document all requirements in Requirements Register (Sheet 2)",
        "3. Assign priorities: Must Have / Should Have / Nice to Have",
        "4. Track implementation status through lifecycle",
        "5. Update Traceability Matrix linking requirements to tests",
        "6. Complete Verification Checklist for testing tracking",
        "7. Conduct Gap Analysis for unimplemented requirements"
    ], 1):
        ws[f"A{row}"] = instr
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        row += 1
    ws.column_dimensions["A"].width = 80

def create_requirements_register(ws, styles):
    ws.merge_cells("A1:M1")
    ws["A1"] = "SECURITY REQUIREMENTS REGISTER\\n13-Field Documentation and Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40
    row = 3
    headers = [("ID", 12), ("Category", 20), ("Requirement Statement", 60), ("Source", 25),
               ("Priority", 15), ("Acceptance Criteria", 40), ("Impl Status", 18),
               ("Assigned To", 20), ("Target Date", 12), ("Verification", 20),
               ("Test Status", 12), ("Evidence", 40), ("Notes", 30)]
    for col_idx, (header, width) in enumerate(headers, 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws[f"{col}{row}"].alignment = styles["column_header"]["alignment"]
        ws[f"{col}{row}"].border = styles["border"]
        ws.column_dimensions[col].width = width
    row += 1
    val_cat = DataValidation(type="list", formula1='"Application Security,Data Protection,Access Control & Authentication,Infrastructure Security,Third-Party Security,Compliance & Regulatory"', allow_blank=False)
    val_pri = DataValidation(type="list", formula1='"Must Have,Should Have,Nice to Have"', allow_blank=False)
    val_impl = DataValidation(type="list", formula1='"Not Started,In Progress,Implemented,Verified"', allow_blank=False)
    val_ver = DataValidation(type="list", formula1='"Functional Test,SAST,DAST,Penetration Test,Vulnerability Scan,Configuration Review,Code Review"', allow_blank=False)
    val_test = DataValidation(type="list", formula1='"Not Tested,Pass,Fail,Blocked,N/A"', allow_blank=False)
    for v in [val_cat, val_pri, val_impl, val_ver, val_test]:
        ws.add_data_validation(v)
    for r in range(row, row + 200):
        ws[f"A{r}"] = f'=TEXT(ROW()-{row-1},"REQ-000")'
        ws[f"A{r}"].font = Font(color="808080")
        for c in range(2, 14):
            col = get_column_letter(c)
            ws[f"{col}{r}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{r}"].border = styles["border"]
        val_cat.add(ws[f"B{r}"])
        val_pri.add(ws[f"E{r}"])
        val_impl.add(ws[f"G{r}"])
        val_ver.add(ws[f"J{r}"])
        val_test.add(ws[f"K{r}"])
        ws[f"I{r}"].number_format = "DD.MM.YYYY"
    ws.freeze_panes = f"A{row}"

def create_example_sheet(ws, styles, category_name, examples):
    ws.merge_cells("A1:C1")
    ws["A1"] = f"EXAMPLE REQUIREMENTS: {category_name.upper()}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    row = 3
    headers = ["Example Requirement", "Acceptance Criteria", "Verification Method"]
    widths = [60, 50, 25]
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws.column_dimensions[col].width = width
    row += 1
    for req, criteria, method in examples:
        ws[f"A{row}"] = req
        ws[f"B{row}"] = criteria
        ws[f"C{row}"] = method
        for col in ['A', 'B', 'C']:
            ws[f"{col}{row}"].alignment = Alignment(wrap_text=True)
            ws[f"{col}{row}"].border = styles["border"]
        row += 1

def create_traceability_matrix(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "TRACEABILITY MATRIX\\nRequirements → Design → Implementation → Tests"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    row = 3
    headers = ["Req ID", "Design Artifact", "Implementation Reference", "Test Case ID", "Test Result", "Verified"]
    widths = [12, 40, 40, 20, 15, 12]
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws.column_dimensions[col].width = width
    row += 1
    val_result = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A"', allow_blank=False)
    val_verified = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(val_result)
    ws.add_data_validation(val_verified)
    for r in range(row, row + 100):
        for c in range(1, 7):
            col = get_column_letter(c)
            ws[f"{col}{r}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{r}"].border = styles["border"]
        val_result.add(ws[f"E{r}"])
        val_verified.add(ws[f"F{r}"])
    ws.freeze_panes = f"A{row}"

def create_verification_checklist(ws, styles):
    ws.merge_cells("A1:D1")
    ws["A1"] = "VERIFICATION CHECKLIST\\nSystematic Testing Status Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    row = 3
    headers = ["Requirement ID", "Verification Method", "Test Status", "Notes"]
    widths = [15, 30, 15, 50]
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws.column_dimensions[col].width = width
    row += 1
    val_status = DataValidation(type="list", formula1='"Not Tested,Pass,Fail,Blocked"', allow_blank=False)
    ws.add_data_validation(val_status)
    for r in range(row, row + 100):
        for c in range(1, 5):
            col = get_column_letter(c)
            ws[f"{col}{r}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{r}"].border = styles["border"]
        val_status.add(ws[f"C{r}"])

def create_gap_analysis(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "GAP ANALYSIS\\nUnimplemented/Failed Requirements with Remediation"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    row = 3
    headers = ["Req ID", "Gap Description", "Impact", "Remediation Action", "Owner", "Target Date"]
    widths = [12, 50, 15, 50, 20, 15]
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws.column_dimensions[col].width = width
    row += 1
    val_impact = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=False)
    ws.add_data_validation(val_impact)
    for r in range(row, row + 50):
        for c in range(1, 7):
            col = get_column_letter(c)
            ws[f"{col}{r}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{r}"].border = styles["border"]
        val_impact.add(ws[f"C{r}"])
        ws[f"F{r}"].number_format = "DD.MM.YYYY"

def create_evidence_register(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "EVIDENCE REGISTER\\nAudit Evidence Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    row = 3
    headers = ["Evidence ID", "Requirement ID", "Description", "Location/Path", "Date Collected", "Status"]
    widths = [12, 12, 50, 45, 15, 18]
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws.column_dimensions[col].width = width
    row += 1
    val_status = DataValidation(type="list", formula1='"Collected,Verified,Pending"', allow_blank=False)
    ws.add_data_validation(val_status)
    for r in range(row, row + 100):
        ws[f"A{r}"] = f'=TEXT(ROW()-{row-1},"EV-000")'
        ws[f"A{r}"].font = Font(color="808080")
        for c in range(2, 7):
            col = get_column_letter(c)
            ws[f"{col}{r}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{r}"].border = styles["border"]
        val_status.add(ws[f"F{r}"])
        ws[f"E{r}"].number_format = "DD.MM.YYYY"
    ws.freeze_panes = f"A{row}"

def create_signoff_sheet(ws, styles):
    ws.merge_cells("A1:E1")
    ws["A1"] = "SIGN-OFF & APPROVAL\\nMulti-Stakeholder Approval"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    row = 3
    ws[f"A{row}"] = "APPROVAL WORKFLOW"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    row += 1
    headers = ["Role", "Name", "Signature", "Date", "Decision"]
    for col_idx, header in enumerate(headers, 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
    row += 1
    val_decision = DataValidation(type="list", formula1='"✅ Approved,⚠️ Conditional,❌ Rejected"', allow_blank=False)
    ws.add_data_validation(val_decision)
    for role in ["Project Manager", "Security Team", "Final Approver"]:
        ws[f"A{row}"] = role
        ws[f"A{row}"].font = Font(bold=True)
        for c in range(2, 6):
            col = get_column_letter(c)
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = styles["border"]
        val_decision.add(ws[f"E{row}"])
        ws[f"D{row}"].number_format = "DD.MM.YYYY"
        row += 1
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20

# Example requirements data
APP_SEC_EXAMPLES = [
    ("System SHALL implement input validation for all user inputs using allowlist approach", "All inputs validated; SQL injection test passes; XSS test passes", "SAST + Penetration Test"),
    ("System SHALL enforce password complexity: minimum 12 characters, uppercase, lowercase, number, special", "Password policy configured; Weak password rejected", "Configuration Review"),
    ("System SHALL implement session timeout after 15 minutes of inactivity", "Session expires after 15 min; User redirected to login", "Functional Test"),
    ("System SHALL implement Role-Based Access Control (RBAC) with principle of least privilege", "All functions require role; Privilege escalation test fails", "Penetration Test"),
    ("System SHALL encrypt all sensitive data in transit using TLS 1.2 or higher", "All endpoints use HTTPS; TLS 1.0/1.1 disabled", "Vulnerability Scan"),
    ("System SHALL hash passwords using bcrypt with work factor ≥12", "Passwords hashed; bcrypt confirmed; Work factor ≥12", "Code Review"),
    ("System SHALL implement proper error handling without exposing sensitive information", "Generic error messages; Stack traces not exposed", "SAST"),
    ("System SHALL log all authentication attempts with timestamp and source IP", "Login events logged; Source IP recorded", "Configuration Review"),
    ("System SHALL implement API rate limiting: 100 requests per minute per API key", "Rate limit enforced; 101st request rejected", "Automated Test"),
    ("System SHALL validate and sanitize all file uploads: type check, size limit 10MB, malware scan", "File type validated; Malware scan performed", "Functional Test")
]

DATA_PROT_EXAMPLES = [
    ("System SHALL encrypt all PII at rest using AES-256", "PII fields encrypted; AES-256 verified", "Configuration Review"),
    ("System SHALL encrypt Restricted data in transit using TLS 1.3", "TLS 1.3 enabled; SSL Labs A+ rating", "Vulnerability Scan"),
    ("System SHALL implement data classification labeling", "All datasets classified; Labels visible", "Data Inspection"),
    ("System SHALL implement automated backup daily with 30-day retention", "Daily backups configured; 30-day retention verified", "Configuration Review"),
    ("System SHALL encrypt all backups using AES-256", "Backup encryption enabled; AES-256 verified", "Configuration Review"),
    ("System SHALL implement point-in-time recovery with 7-day window", "PITR configured; 7-day window available", "Functional Test"),
    ("System SHALL implement data retention: Restricted 7 years, Confidential 3 years", "Retention policy documented; Automated deletion configured", "Document Review"),
    ("System SHALL implement secure data deletion using crypto-erasure or overwriting", "Deletion method documented; Deleted data unrecoverable", "Functional Test"),
    ("System SHALL implement data minimization: collect only necessary data", "Data collection justified; Unnecessary fields not collected", "Document Review"),
    ("System SHALL implement DLP controls preventing unauthorized data exfiltration", "DLP solution deployed; Restricted data egress blocked", "Configuration Review")
]

ACCESS_CTRL_EXAMPLES = [
    ("System SHALL implement least privilege access control", "Users have minimum required permissions", "Penetration Test"),
    ("System SHALL implement multi-factor authentication for privileged accounts", "Admin accounts require MFA; Password-only blocked", "Configuration Review"),
    ("System SHALL implement account lifecycle management", "Account creation/modification/deletion documented", "Document Review"),
    ("System SHALL implement privileged access management", "Privileged access logged and monitored", "Configuration Review"),
    ("System SHALL implement password rotation policy: 90 days maximum", "Password rotation enforced; 90-day max verified", "Configuration Review")
]

INFRA_EXAMPLES = [
    ("System SHALL implement network segmentation separating prod from dev", "Network zones documented; Firewall enforcing segmentation", "Configuration Review"),
    ("System SHALL implement firewall rules following deny-by-default", "Default deny configured; Only required ports open", "Configuration Review"),
    ("System SHALL implement server hardening per CIS benchmarks", "CIS benchmark applied; Hardening verified", "Vulnerability Scan"),
    ("System SHALL implement patch management: critical patches within 30 days", "Patch management process documented; Critical patches applied", "Configuration Review"),
    ("System SHALL implement security monitoring and SIEM integration", "SIEM configured; Security events forwarded", "Configuration Review")
]

THIRD_PARTY_EXAMPLES = [
    ("System SHALL conduct security assessment of vendors handling Restricted data", "Vendor security questionnaire completed; Assessment documented", "Document Review"),
    ("System SHALL implement vendor contracts with security requirements", "Contracts include security clauses; NDA executed", "Document Review"),
    ("System SHALL conduct API security assessment for third-party integrations", "API security review completed; Vulnerabilities remediated", "Penetration Test")
]

COMPLIANCE_EXAMPLES = [
    ("System SHALL comply with GDPR Article 32 security requirements", "TOMs documented; Risk assessment completed", "Document Review"),
    ("System SHALL comply with nDSG data protection requirements", "Swiss data residency verified; DPA executed", "Document Review"),
    ("System SHALL implement audit logging per ISO 27001 requirements", "Audit logs configured; 6-month retention", "Configuration Review")
]

def main():
    print("=" * 78)
    print(f"{DOCUMENT_ID} - Security Requirements Register Generator")
    print(CONTROL_REF)
    print("=" * 78)
    wb = create_workbook()
    styles = setup_styles()
    print("\\n[1/13] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    print("[2/13] Creating Requirements Register (200 rows)...")
    create_requirements_register(wb["Requirements Register"], styles)
    print("[3/13] Creating App Security Examples...")
    create_example_sheet(wb["App Security Examples"], styles, "Application Security", APP_SEC_EXAMPLES)
    print("[4/13] Creating Data Protection Examples...")
    create_example_sheet(wb["Data Protection Examples"], styles, "Data Protection", DATA_PROT_EXAMPLES)
    print("[5/13] Creating Access Control Examples...")
    create_example_sheet(wb["Access Control Examples"], styles, "Access Control & Authentication", ACCESS_CTRL_EXAMPLES)
    print("[6/13] Creating Infrastructure Examples...")
    create_example_sheet(wb["Infrastructure Examples"], styles, "Infrastructure Security", INFRA_EXAMPLES)
    print("[7/13] Creating Third-Party Examples...")
    create_example_sheet(wb["Third-Party Examples"], styles, "Third-Party Security", THIRD_PARTY_EXAMPLES)
    print("[8/13] Creating Compliance Examples...")
    create_example_sheet(wb["Compliance Examples"], styles, "Compliance & Regulatory", COMPLIANCE_EXAMPLES)
    print("[9/13] Creating Traceability Matrix...")
    create_traceability_matrix(wb["Traceability Matrix"], styles)
    print("[10/13] Creating Verification Checklist...")
    create_verification_checklist(wb["Verification Checklist"], styles)
    print("[11/13] Creating Gap Analysis...")
    create_gap_analysis(wb["Gap Analysis"], styles)
    print("[12/13] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)
    print("[13/13] Creating Sign-Off...")
    create_signoff_sheet(wb["Sign-Off"], styles)
    filename = f"{DOCUMENT_ID}_Security_Requirements_Register_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"\\n✅ SUCCESS: {filename}")
    print(f"📄 File: {filename}")
    print(f"📊 Sheets: 13 sheets created")
    print(f"📝 Requirements: 200 pre-formatted rows")
    print(f"📋 Examples: 10-10-5-5-3-3 requirements per category")
    print("=" * 78)

if __name__ == "__main__":
    main()
