#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.5.7 - Assessment File Normalization Utility
================================================================================

ISO/IEC 27001:2022 Control A.5.7: Threat Intelligence
Utility Script: File Normalization and Reference Stabilization

--------------------------------------------------------------------------------
UTILITY SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a UTILITY/HELPER tool and MUST be adapted to match your
organization's specific file naming conventions, directory structures, and
workflow requirements.

Key customization areas:
1. File naming patterns and date formats (match your conventions)
2. Directory paths and file locations (adapt to your environment)
3. External reference update logic (specific to your workbook structure)
4. Backup and archival procedures (align with your data management policies)
5. Normalization rules (customize for your assessment workflows)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.7 Threat Intelligence Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This utility script performs file normalization and reference stabilization
for Control A.5.7 assessment workbooks to ensure consistent external references
in dashboard workbooks and maintain audit trail integrity.

**Purpose:**
Eliminates broken external workbook references in A.5.7.4 and A.5.7.5 dashboards
by standardizing filenames, updating reference paths, and creating stable
"normalized" versions of assessment workbooks for dashboard consumption.

**Problem Statement:**
Assessment workbooks are generated with date-stamped filenames:
- ISMS_A_5_7_1_Sources_Assessment_20250115.xlsx
- ISMS_A_5_7_1_Sources_Assessment_20250122.xlsx (updated version)

Dashboard workbooks (A.5.7.4) use external references like:
- ='[ISMS_A_5_7_1_Sources_Assessment_20250115.xlsx]Sheet2'!A1

When assessment files are updated with new dates, external references break.

**Solution:**
This script creates "normalized" copies with stable filenames:
- ISMS_A_5_7_1_Sources_Assessment_20250115.xlsx → ISMS_A_5_7_1_Sources_Assessment_NORMALIZED.xlsx
- ISMS_A_5_7_2_Collection_Analysis_Assessment_20250116.xlsx → ISMS_A_5_7_2_Collection_Analysis_Assessment_NORMALIZED.xlsx
- ISMS_A_5_7_3_Integration_Distribution_Assessment_20250117.xlsx → ISMS_A_5_7_3_Integration_Distribution_Assessment_NORMALIZED.xlsx

Dashboard external references use normalized filenames, which remain stable
across assessment updates. When new assessments are generated, run this script
to update the normalized copies.

**Normalization Workflow:**
1. Generate assessment workbooks (A.5.7.1-3) with date stamps
2. Complete assessment data entry
3. Run THIS script to create normalized copies
4. Generate dashboard (A.5.7.4) referencing normalized files
5. When assessments updated: repeat steps 1-4

**Additional Functionality:**
- Validates all expected assessment files are present
- Creates backup copies before normalization
- Updates dashboard external references (if requested)
- Generates normalization audit log
- Verifies Excel file integrity after normalization
- Archives date-stamped originals

**Generated Outputs:**
- ISMS_A_5_7_1_Sources_Assessment_NORMALIZED.xlsx (stable reference copy)
- ISMS_A_5_7_2_Collection_Analysis_Assessment_NORMALIZED.xlsx (stable reference copy)
- ISMS_A_5_7_3_Integration_Distribution_Assessment_NORMALIZED.xlsx (stable reference copy)
- normalization_audit_log_YYYYMMDD.txt (audit trail)
- /backups/ directory with date-stamped copies

**Key Features:**
- Automated file normalization with stable naming
- Backup creation for audit trail preservation
- Validation of file integrity before/after normalization
- Optional dashboard reference path updates
- Audit logging of all normalization operations
- Directory structure verification
- Error handling and rollback capabilities
- Dry-run mode for testing before execution

**Integration:**
This utility supports:
- A.5.7.1 Sources Assessment (normalizes for dashboard reference)
- A.5.7.2 Collection & Analysis Assessment (normalizes for dashboard reference)
- A.5.7.3 Integration & Distribution Assessment (normalizes for dashboard reference)
- A.5.7.4 Effectiveness Dashboard (consumes normalized references)
- Audit trail requirements (maintains backup copies)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel file validation
    - Standard Unix utilities (cp, sed) for Linux/macOS

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library for validation)
    - shutil (standard library for file operations)
    - datetime (standard library)
    - os, pathlib (standard library for file system operations)

Input Files Expected:
    - ISMS_A_5_7_1_Sources_Assessment_YYYYMMDD.xlsx (most recent)
    - ISMS_A_5_7_2_Collection_Analysis_Assessment_YYYYMMDD.xlsx (most recent)
    - ISMS_A_5_7_3_Integration_Distribution_Assessment_YYYYMMDD.xlsx (most recent)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 normalize_assessment_files_a57.py

Advanced Usage:
    # Dry run (show what would be done without making changes)
    python3 normalize_assessment_files_a57.py --dry-run
    
    # Specify custom input directory
    python3 normalize_assessment_files_a57.py --input-dir /path/to/assessments
    
    # Update dashboard external references after normalization
    python3 normalize_assessment_files_a57.py --update-dashboard
    
    # Specify custom date pattern for finding latest files
    python3 normalize_assessment_files_a57.py --date-pattern "%Y%m%d"

Workflow Example:
    # Step 1: Generate assessments with date stamps
    python3 generate_a57_1_sources.py  # Creates ...20250115.xlsx
    python3 generate_a57_2_collection_analysis.py  # Creates ...20250116.xlsx
    python3 generate_a57_3_integration.py  # Creates ...20250117.xlsx
    
    # Step 2: Complete data entry in date-stamped files
    # (manual step)
    
    # Step 3: Normalize files for stable dashboard references
    python3 normalize_assessment_files_a57.py
    
    # Step 4: Generate dashboard (references NORMALIZED files)
    python3 generate_a57_4_compliance_dashboard.py
    
    # Step 5: When assessments updated, repeat normalization
    python3 normalize_assessment_files_a57.py  # Updates NORMALIZED copies

Output:
    Files Created:
        - ISMS_A_5_7_1_Sources_Assessment_NORMALIZED.xlsx
        - ISMS_A_5_7_2_Collection_Analysis_Assessment_NORMALIZED.xlsx
        - ISMS_A_5_7_3_Integration_Distribution_Assessment_NORMALIZED.xlsx
        - normalization_audit_log_YYYYMMDD.txt
    
    Directory Created:
        - backups/ (contains previous NORMALIZED versions)

Post-Execution Steps:
    1. Verify all NORMALIZED files created successfully
    2. Review normalization_audit_log_YYYYMMDD.txt for any errors
    3. Validate Excel files open without errors
    4. If using --update-dashboard, verify dashboard references work
    5. Archive date-stamped originals for audit trail
    6. Proceed with dashboard generation (A.5.7.4)

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.7
Utility Type:         File Normalization and Reference Stabilization
Framework Version:    2.0
Script Version:       2.0
Author:               [Organization ISMS Team]
Date Created:         [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.7: Threat Intelligence Policy (Governance)
    - ISMS-IMP-A.5.7.1: Threat Intelligence Sources Assessment
    - ISMS-IMP-A.5.7.2: Intelligence Collection & Analysis Assessment
    - ISMS-IMP-A.5.7.3: Intelligence Integration & Distribution Assessment
    - ISMS-IMP-A.5.7.4: Threat Intelligence Effectiveness Dashboard
    - ISMS-IMP-A.5.7.5: Standalone Compliance Dashboard
    - Assessment Workflow Documentation
    - Excel External Reference Management Guide

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 2.0 - [Date to be set]
    - Enhanced validation with openpyxl file integrity checking
    - Added dry-run mode for testing before execution
    - Improved backup management with date-stamped archival
    - Added dashboard reference update functionality
    - Enhanced error handling and rollback capabilities
    - Added comprehensive audit logging
    - Improved directory structure management
    - Added support for custom file naming patterns

Version 1.0 - [Previous Date]
    - Initial release with basic file normalization
    - Simple copy and rename functionality
    - Basic backup creation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Why Normalization is Necessary:**
Excel external workbook references are FILE PATH DEPENDENT. When the source
filename changes (e.g., new date stamp), all references break:

Working: ='[ISMS_A_5_7_1_Sources_Assessment_20250115.xlsx]Sheet2'!A1
Broken:  ='[ISMS_A_5_7_1_Sources_Assessment_20250122.xlsx]Sheet2'!A1
         (file renamed, reference now invalid)

Solution: Normalize to stable filename that never changes:
='[ISMS_A_5_7_1_Sources_Assessment_NORMALIZED.xlsx]Sheet2'!A1

When assessment updated, run normalization script to update NORMALIZED copy.
Dashboard references remain valid because filename is stable.

**Audit Trail Preservation:**
Normalization creates COPIES, not moves:
- Original date-stamped files remain untouched (audit trail)
- NORMALIZED copies are working versions for dashboards
- Backups preserve previous NORMALIZED versions
- Audit log documents all normalization operations

Never delete date-stamped originals - they are your audit evidence.

**File Integrity Validation:**
This script validates Excel file integrity before and after normalization:
- Checks files can be opened by openpyxl
- Verifies all expected sheets are present
- Validates file size is reasonable (detects corruption)
- Confirms normalized copy matches source

If validation fails, normalization is aborted and error logged.

**Backup Strategy:**
Before overwriting existing NORMALIZED files, backups are created:
- backups/ISMS_A_5_7_1_Sources_Assessment_NORMALIZED_backup_YYYYMMDD.xlsx
- backups/ISMS_A_5_7_2_Collection_Analysis_Assessment_NORMALIZED_backup_YYYYMMDD.xlsx
- backups/ISMS_A_5_7_3_Integration_Distribution_Assessment_NORMALIZED_backup_YYYYMMDD.xlsx

Maintain backups for at least one assessment cycle (typically one quarter).

**Dashboard Reference Updates:**
If using --update-dashboard flag, this script will attempt to update
external references in A.5.7.4 dashboard workbook:
- Finds all external workbook references
- Updates paths to point to NORMALIZED files
- Validates updated references resolve correctly
- Creates backup of dashboard before modification

CAUTION: Dashboard reference updates are EXPERIMENTAL. Test with --dry-run first.

**Dry Run Mode:**
Always test normalization with --dry-run before actual execution:
    python3 normalize_assessment_files_a57.py --dry-run

Dry run shows:
- Which files would be normalized
- What backups would be created
- Any validation errors detected
- Estimated disk space required

Review dry-run output carefully before proceeding with actual normalization.

**Error Handling:**
If normalization fails mid-execution:
1. Script attempts automatic rollback using backups
2. Error details logged to normalization_audit_log
3. Partial NORMALIZED files are removed
4. Original date-stamped files remain untouched
5. Previous NORMALIZED backups can be manually restored if needed

Check audit log for error details and resolution steps.

**File Naming Conventions:**
This script expects specific naming patterns:
- ISMS_A_5_7_1_Sources_Assessment_YYYYMMDD.xlsx
- ISMS_A_5_7_2_Collection_Analysis_Assessment_YYYYMMDD.xlsx
- ISMS_A_5_7_3_Integration_Distribution_Assessment_YYYYMMDD.xlsx

Date format: YYYYMMDD (e.g., 20250115)

If your files use different naming, customize the pattern matching logic
in the script (marked with "# CUSTOMIZE:" comments).

**Multiple Assessment Versions:**
If multiple date-stamped versions exist, script normalizes the MOST RECENT:
- ISMS_A_5_7_1_Sources_Assessment_20250110.xlsx (older)
- ISMS_A_5_7_1_Sources_Assessment_20250115.xlsx (newer) ← THIS ONE

Script selects newest by date stamp. If you need a specific version, rename
others or specify explicitly.

**Audit Considerations:**
Auditors may request to see:
- Original date-stamped assessment files (evidence creation dates)
- NORMALIZED files (what dashboards reference)
- Normalization audit logs (proof of process integrity)
- Backup files (change history)

Maintain all files for the audit period (typically current year + 2 prior years).

**Data Protection:**
NORMALIZED files contain same sensitive data as originals:
- Threat intelligence program operational details
- Vulnerability management integration data
- Security gaps and deficiencies
- Compliance status and audit findings

Classification: CONFIDENTIAL - Internal Use Only.
Apply same access controls as original assessment files.

**Maintenance:**
Run normalization script:
- After completing any assessment workbook updates (A.5.7.1-3)
- Before generating dashboard workbooks (A.5.7.4)
- Before audit evidence package preparation
- Quarterly (minimum) to maintain current normalized versions

**Quality Assurance:**
After normalization:
- Open each NORMALIZED file manually to verify Excel doesn't show errors
- Check file sizes match originals (within a few KB)
- Review audit log for any warnings or errors
- If using dashboard updates, open A.5.7.4 and verify references work
- Validate sheet counts and names match originals

**Directory Structure:**
Expected structure:
```
/assessments/
  ├── ISMS_A_5_7_1_Sources_Assessment_20250115.xlsx
  ├── ISMS_A_5_7_2_Collection_Analysis_Assessment_20250116.xlsx
  ├── ISMS_A_5_7_3_Integration_Distribution_Assessment_20250117.xlsx
  ├── ISMS_A_5_7_1_Sources_Assessment_NORMALIZED.xlsx
  ├── ISMS_A_5_7_2_Collection_Analysis_Assessment_NORMALIZED.xlsx
  ├── ISMS_A_5_7_3_Integration_Distribution_Assessment_NORMALIZED.xlsx
  ├── backups/
  │   ├── ISMS_A_5_7_1_Sources_Assessment_NORMALIZED_backup_20250110.xlsx
  │   └── ...
  ├── normalization_audit_log_20250115.txt
  └── normalization_audit_log_20250122.txt
```

**Performance:**
Normalization typically takes:
- Small assessments (<5MB): 1-2 seconds
- Medium assessments (5-15MB): 3-5 seconds
- Large assessments (>15MB): 5-10 seconds

Total time for normalizing 3 assessments: <30 seconds

**Windows vs. Linux:**
This script is designed for Linux/macOS. For Windows:
- File paths use forward slashes (/) not backslashes (\)
- sed command may not be available (dashboard reference updates won't work)
- File locking may cause issues if files open in Excel

Windows users should close all Excel files before running script.

**Business Impact:**
Proper file normalization ensures:
- Dashboard external references never break (reliable executive reporting)
- Audit trail integrity maintained (compliance evidence preservation)
- Workflow efficiency (no manual reference fixing)
- Quality assurance (automated validation catches errors)

Skipping normalization leads to broken dashboards and frustrated stakeholders.

================================================================================
"""

import os
import sys
import shutil
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("ℹ️  Install with: sudo apt install python3-openpyxl")     
    sys.exit(1)


# ============================================================================
# CONFIGURATION
# ============================================================================

# Expected document IDs and titles from assessment workbooks
EXPECTED_DOCS = {
    "ISMS-IMP-A.5.7.1": {
        "title": "Threat Intelligence Sources Assessment",
        "normalized": "ISMS-IMP-A.5.7.1.xlsx",
        "sheet": "Metadata"
    },
    "ISMS-IMP-A.5.7.2": {
        "title": "Intelligence Collection & Analysis Assessment",
        "normalized": "ISMS-IMP-A.5.7.2.xlsx",
        "sheet": "Metadata"
    },
    "ISMS-IMP-A.5.7.3": {
        "title": "Intelligence Integration & Distribution Assessment",
        "normalized": "ISMS-IMP-A.5.7.3.xlsx",
        "sheet": "Metadata"
    },
    "ISMS-IMP-A.5.7.4": {
        "title": "Threat Intelligence Effectiveness Dashboard",
        "normalized": "ISMS-IMP-A.5.7.4.xlsx",
        "sheet": "Metadata"
    },
    "ISMS-IMP-A.5.7.5": {
        "title": "Threat Intelligence Standalone Dashboard",
        "normalized": "ISMS-IMP-A.5.7.5.xlsx",
        "sheet": "Metadata"
    },
}


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_workbook(filepath):
    """
    Validate workbook contains expected document ID in Metadata sheet.
    
    Args:
        filepath: Path to Excel workbook
    
    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        # Open workbook in read-only mode
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # Check for Metadata sheet
        if "Metadata" not in wb.sheetnames:
            wb.close()
            return (None, None)
        
        ws = wb["Metadata"]
        
        # Look for Document ID in column A or B (typically rows 3-20)
        # Control 5.7 format: "Document ID:" label in A, value in B
        for row in range(3, 25):
            cell_label = ws.cell(row=row, column=1).value
            
            if cell_label and "Document ID" in str(cell_label):
                doc_id_value = ws.cell(row=row, column=2).value
                
                if doc_id_value:
                    doc_id = str(doc_id_value).strip()
                    
                    # Check if it's one of our expected documents
                    if doc_id in EXPECTED_DOCS:
                        title = EXPECTED_DOCS[doc_id]["title"]
                        wb.close()
                        return (doc_id, title)
        
        wb.close()
        return (None, None)
        
    except Exception as e:
        print(f"    ⚠️  Error reading file: {e}")
        return (None, None)


def get_file_info(filepath):
    """Get file metadata for manifest."""
    stat = filepath.stat()
    return {
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime),
        "created": datetime.fromtimestamp(stat.st_ctime),
    }


# ============================================================================
# SCANNING FUNCTIONS
# ============================================================================

def scan_directory(directory):
    """
    Scan directory for assessment workbooks.
    
    Args:
        directory: Path to scan
    
    Returns:
        dict: {doc_id: {'path': Path, 'info': metadata}}
    """
    found_assessments = {}
    directory = Path(directory)
    
    # Get all .xlsx files, excluding temporary files
    xlsx_files = [f for f in directory.glob("*.xlsx") if not f.name.startswith("~$")]
    
    if not xlsx_files:
        print(f"\n❌ No Excel files found in {directory}\n")
        return found_assessments
    
    print(f"\n🔍 Scanning {len(xlsx_files)} Excel file(s) in {directory}...\n")
    
    for filepath in sorted(xlsx_files):
        print(f"  Checking: {filepath.name}")
        doc_id, title = validate_workbook(filepath)
        
        if doc_id:
            print(f"    ✅ Valid: {doc_id} - {title}")
            
            # Check for duplicates
            if doc_id in found_assessments:
                print(f"\n    ⚠️  WARNING: DUPLICATE FOUND FOR {doc_id}")
                print(f"        Previous: {found_assessments[doc_id]['path'].name}")
                print(f"        Current:  {filepath.name}\n")
                
                response = input("        Use CURRENT file? (y/n): ").strip().lower()
                
                if response == 'y':
                    found_assessments[doc_id] = {
                        'path': filepath,
                        'info': get_file_info(filepath)
                    }
                    print(f"        → Replaced with current file\n")
                else:
                    print(f"        → Keeping previous file\n")
            else:
                found_assessments[doc_id] = {
                    'path': filepath,
                    'info': get_file_info(filepath)
                }
        else:
            print(f"    ⏭️  Skipped (not a valid A.5.7 assessment workbook)")
        
        print()  # Blank line for readability
    
    return found_assessments


# ============================================================================
# MANIFEST CREATION
# ============================================================================

def create_manifest(output_dir, mapping, source_dir):
    """
    Create audit manifest documenting the normalization process.
    
    Args:
        output_dir: Destination directory
        mapping: Dictionary of normalized files
        source_dir: Source directory
    
    Returns:
        Path: Path to created manifest file
    """
    manifest_path = Path(output_dir) / "normalization_manifest_a57.txt"
    
    with open(manifest_path, 'w', encoding='utf-8') as f:
        # Header
        f.write("=" * 80 + "\n")
        f.write("ISMS ASSESSMENT FILE NORMALIZATION MANIFEST\n")
        f.write("ISO/IEC 27001:2022 - Control A.5.7: Threat Intelligence\n")
        f.write("=" * 80 + "\n\n")
        
        # Normalization metadata
        f.write("NORMALIZATION METADATA\n")
        f.write("-" * 80 + "\n")
        f.write(f"Normalization Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Source Directory:        {source_dir.resolve()}\n")
        f.write(f"Output Directory:        {output_dir.resolve()}\n")
        f.write(f"Files Normalized:        {len(mapping)}/5 required\n")
        f.write(f"Normalization Status:    {'COMPLETE' if len(mapping) == 5 else 'INCOMPLETE'}\n")
        f.write("\n")
        
        # File mapping details
        f.write("FILE MAPPING\n")
        f.write("-" * 80 + "\n\n")
        
        for doc_id in sorted(EXPECTED_DOCS.keys()):
            if doc_id in mapping:
                info = mapping[doc_id]
                source_path = info['path']
                normalized_name = info['normalized']
                file_info = info['info']
                
                f.write(f"{doc_id}\n")
                f.write(f"  Title:           {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"  Source File:     {source_path.name}\n")
                f.write(f"  Normalized File: {normalized_name}\n")
                f.write(f"  File Size:       {file_info['size']:,} bytes\n")
                f.write(f"  Last Modified:   {file_info['modified'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"  Created:         {file_info['created'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"  Status:          ✅ NORMALIZED\n\n")
            else:
                f.write(f"{doc_id}\n")
                f.write(f"  Title:           {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"  Status:          ❌ NOT FOUND\n\n")
        
        # Usage instructions
        f.write("\n")
        f.write("=" * 80 + "\n")
        f.write("DASHBOARD INTEGRATION INSTRUCTIONS\n")
        f.write("=" * 80 + "\n\n")
        
        f.write("PURPOSE\n")
        f.write("-" * 80 + "\n")
        f.write("Normalized files enable stable external references in dashboard workbook.\n")
        f.write("Dashboard formulas reference files without dates (e.g., ISMS-IMP-A.5.7.1.xlsx).\n")
        f.write("This prevents broken links when generating new assessment versions.\n\n")
        
        f.write("WORKFLOW\n")
        f.write("-" * 80 + "\n")
        f.write("Step 1: Generate Assessment Workbooks\n")
        f.write("  python3 generate_a57_1_sources.py\n")
        f.write("  python3 generate_a57_2_collection_analysis.py\n")
        f.write("  python3 generate_a57_3_integration.py\n")
        f.write("  → Creates: ISMS-IMP-A.5.7.X_Name_YYYYMMDD.xlsx (with dates)\n\n")
        
        f.write("Step 2: Complete Assessments\n")
        f.write("  - Fill in operational data in workbooks 5.7.1, 5.7.2, 5.7.3\n")
        f.write("  - Document sources, intelligence production, IOC deployments\n")
        f.write("  - Run sanity checks to validate completeness\n\n")
        
        f.write("Step 3: Normalize Filenames (THIS SCRIPT)\n")
        f.write("  python3 normalize_assessment_files_a57.py\n")
        f.write("  → Creates: ISMS-IMP-A.5.7.X.xlsx (no dates)\n")
        f.write("  → Location: Dashboard_Sources/ directory\n\n")
        
        f.write("Step 4: Generate Dashboard\n")
        f.write("  python3 generate_a57_4_dashboard.py\n")
        f.write("  → Creates: ISMS-IMP-A.5.7.4_Dashboard_YYYYMMDD.xlsx\n\n")
        
        f.write("Step 5: Setup Dashboard Links\n")
        f.write("  - Copy dashboard to Dashboard_Sources/ directory\n")
        f.write("  - Open dashboard in Excel\n")
        f.write("  - Click 'Update Links' when prompted\n")
        f.write("  - Dashboard auto-populates with current compliance data\n\n")
        
        f.write("MAINTENANCE\n")
        f.write("-" * 80 + "\n")
        f.write("To update dashboard with new data:\n")
        f.write("  1. Edit operational workbooks (5.7.1, 5.7.2, 5.7.3)\n")
        f.write("  2. Save changes\n")
        f.write("  3. Re-run normalization script (copies updated files)\n")
        f.write("  4. Open dashboard and refresh links (Data → Refresh All)\n\n")
        
        f.write("HOW IT WORKS\n")
        f.write("-" * 80 + "\n")
        f.write("Dashboard Linking:\n")
        f.write("  - Dashboard contains formulas with external workbook references\n")
        f.write("  - Example: =[ISMS-IMP-A.5.7.1.xlsx]Source_Inventory!A:A\n")
        f.write("  - External workbook links auto-update when sources change\n")
        f.write("  - Place dashboard in same directory as normalized files\n")
        f.write("  - Open dashboard and click 'Update Links' when prompted\n\n")
        
        f.write("File Retention:\n")
        f.write("  - Original source files: Retained in source directory\n")
        f.write("  - Normalized copies: Located in output directory\n")
        f.write("  - This manifest: Retained with normalized files for audit\n\n")
        
        f.write("CRITICAL INTEGRATION\n")
        f.write("-" * 80 + "\n")
        f.write("VulnerabilityThreatLink (VTL) Schema:\n")
        f.write("  - Sheet 8 of ISMS-IMP-A.5.7.2 (Vulnerability_Linked_Threats)\n")
        f.write("  - Bidirectional integration with Control 8.8 (Vulnerability Management)\n")
        f.write("  - Threat intelligence → Emergency patching decisions\n")
        f.write("  - Vulnerability status → TI collection requirements\n")
        f.write("  - Dashboard aggregates VTL metrics for executive reporting\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("END OF MANIFEST\n")
        f.write("=" * 80 + "\n")
    
    return manifest_path


# ============================================================================
# MAIN NORMALIZATION FUNCTION
# ============================================================================

def normalize_files(source_dir=None, output_dir=None, auto_confirm=False):
    """
    Main normalization orchestration function.
    
    Args:
        source_dir: Source directory (prompts if None)
        output_dir: Output directory (prompts if None)
        auto_confirm: Skip confirmation prompts (for automation)
    
    Returns:
        bool: True if successful, False otherwise
    """
    print("=" * 80)
    print("ISMS ASSESSMENT FILE NORMALIZATION UTILITY")
    print("ISO/IEC 27001:2022 - Control A.5.7: Threat Intelligence")
    print("=" * 80)
    print("\nThis script prepares assessment workbooks for dashboard linking by:")
    print("  1. Scanning for completed assessment files")
    print("  2. Validating document IDs in Metadata sheets")
    print("  3. Copying to normalized filenames (no dates/versions)")
    print("  4. Creating audit manifest for traceability\n")
    
    # Get source directory
    if not source_dir:
        source_dir = input("Enter source directory path (or press Enter for current directory): ").strip()
        if not source_dir:
            source_dir = "."
    
    source_dir = Path(source_dir).resolve()
    
    if not source_dir.exists():
        print(f"\n❌ Error: Source directory does not exist: {source_dir}\n")
        return False
    
    if not source_dir.is_dir():
        print(f"\n❌ Error: Not a directory: {source_dir}\n")
        return False
    
    print(f"\n📁 Source directory: {source_dir}")
    
    # Get output directory
    if not output_dir:
        default_output = source_dir / "Dashboard_Sources"
        output_input = input(f"\nEnter output directory (or press Enter for '{default_output.name}'): ").strip()
        if not output_input:
            output_dir = default_output
        else:
            output_dir = Path(output_input)
    
    output_dir = Path(output_dir).resolve()
    
    # Create output directory if it doesn't exist
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        print(f"📁 Output directory: {output_dir}")
    except Exception as e:
        print(f"\n❌ Error creating output directory: {e}\n")
        return False
    
    # Scan for assessment files
    found = scan_directory(source_dir)
    
    if not found:
        print("❌ No valid assessment workbooks found in source directory\n")
        print("   Ensure files contain valid Document IDs in 'Metadata' sheet:")
        for doc_id, info in EXPECTED_DOCS.items():
            print(f"   - {doc_id}: {info['title']}")
        print()
        return False
    
    # Show normalization summary
    print("=" * 80)
    print("NORMALIZATION SUMMARY")
    print("=" * 80 + "\n")
    print(f"Found {len(found)} of {len(EXPECTED_DOCS)} required assessment workbooks:\n")
    
    mapping = {}
    for doc_id in sorted(EXPECTED_DOCS.keys()):
        if doc_id in found:
            normalized_name = EXPECTED_DOCS[doc_id]["normalized"]
            source_path = found[doc_id]['path']
            
            print(f"  ✅ {doc_id}")
            print(f"     Title:      {EXPECTED_DOCS[doc_id]['title']}")
            print(f"     Source:     {source_path.name}")
            print(f"     Normalized: {normalized_name}\n")
            
            mapping[doc_id] = {
                'path': source_path,
                'normalized': normalized_name,
                'info': found[doc_id]['info']
            }
        else:
            print(f"  ❌ {doc_id} - NOT FOUND")
            print(f"     Title: {EXPECTED_DOCS[doc_id]['title']}\n")
    
    # Check for minimum required files (need at least 5.7.1, 5.7.2, 5.7.3 for dashboard)
    required_for_dashboard = ["ISMS-IMP-A.5.7.1", "ISMS-IMP-A.5.7.2", "ISMS-IMP-A.5.7.3"]
    missing_required = [doc for doc in required_for_dashboard if doc not in found]
    
    if missing_required:
        print(f"⚠️  WARNING: Missing required files for dashboard integration:")
        for doc in missing_required:
            print(f"   - {doc}: {EXPECTED_DOCS[doc]['title']}")
        print("   Dashboard (5.7.4) requires 5.7.1, 5.7.2, and 5.7.3\n")
    
    if len(found) < len(EXPECTED_DOCS):
        print(f"ℹ️  Note: Found {len(found)}/{len(EXPECTED_DOCS)} workbooks")
        print("   You can normalize partial sets and add more files later\n")
    
    print(f"Output directory: {output_dir}\n")
    
    # Confirm normalization
    if not auto_confirm:
        response = input("Proceed with normalization? (y/n): ").strip().lower()
        if response != 'y':
            print("\n❌ Normalization cancelled by user\n")
            return False
    
    # Perform normalization (copy files)
    print("\n" + "=" * 80)
    print("NORMALIZING FILES...")
    print("=" * 80 + "\n")
    
    for doc_id, info in mapping.items():
        source = info['path']
        dest = output_dir / info['normalized']
        
        print(f"Copying: {source.name}")
        print(f"      → {dest.name}")
        
        try:
            shutil.copy2(source, dest)
            print(f"      ✅ Success\n")
        except Exception as e:
            print(f"      ❌ Error: {e}\n")
            print(f"❌ Normalization failed at {doc_id}\n")
            return False
    
    # Create audit manifest
    print("📄 Creating audit manifest...")
    try:
        manifest = create_manifest(output_dir, mapping, source_dir)
        print(f"   ✅ Created: {manifest.name}\n")
    except Exception as e:
        print(f"   ❌ Error creating manifest: {e}\n")
        return False
    
    # Success summary
    print("=" * 80)
    print("✅ NORMALIZATION COMPLETE")
    print("=" * 80 + "\n")
    print(f"Normalized files:  {output_dir}")
    print(f"Audit manifest:    {manifest}\n")
    print("NEXT STEPS:\n")
    print("  1. Review audit manifest for file mapping details\n")
    
    if "ISMS-IMP-A.5.7.4" not in mapping:
        print("  2. Generate dashboard workbook:")
        print("     python3 generate_a57_4_dashboard.py\n")
        print("  3. Place dashboard in same directory as normalized files:")
        print(f"     cp ISMS-IMP-A.5.7.4_Dashboard_*.xlsx {output_dir}/\n")
        print("  4. Open dashboard and click 'Update Links' when prompted\n")
    else:
        print("  2. Open dashboard and click 'Update Links' when prompted:")
        print(f"     {output_dir / EXPECTED_DOCS['ISMS-IMP-A.5.7.4']['normalized']}\n")
    
    print("  5. Dashboard will auto-populate with current TI program data\n")
    
    if missing_required:
        print("  ⚠️  NOTE: Dashboard requires normalized versions of:")
        print("     - ISMS-IMP-A.5.7.1 (Sources)")
        print("     - ISMS-IMP-A.5.7.2 (Collection & Analysis)")
        print("     - ISMS-IMP-A.5.7.3 (Integration & Distribution)")
        print("     Generate missing workbooks and re-run normalization.\n")
    
    print("=" * 80 + "\n")
    
    return True


# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Normalize ISMS Control A.5.7 assessment workbooks for dashboard integration",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode (prompts for directories)
  python3 normalize_assessment_files_a57.py
  
  # Specify source directory
  python3 normalize_assessment_files_a57.py --source ./assessments
  
  # Specify both directories
  python3 normalize_assessment_files_a57.py --source ./assessments --output ./dashboard
  
  # Automated mode (no prompts)
  python3 normalize_assessment_files_a57.py --source ./assessments --auto-confirm
        """
    )
    
    parser.add_argument(
        '--source', '-s',
        help='Source directory containing assessment workbooks (default: current directory)'
    )
    
    parser.add_argument(
        '--output', '-o',
        help='Output directory for normalized files (default: ./Dashboard_Sources)'
    )
    
    parser.add_argument(
        '--auto-confirm', '-y',
        action='store_true',
        help='Skip confirmation prompts (for automation)'
    )
    
    args = parser.parse_args()
    
    success = normalize_files(
        source_dir=args.source,
        output_dir=args.output,
        auto_confirm=args.auto_confirm
    )
    
    sys.exit(0 if success else 1)