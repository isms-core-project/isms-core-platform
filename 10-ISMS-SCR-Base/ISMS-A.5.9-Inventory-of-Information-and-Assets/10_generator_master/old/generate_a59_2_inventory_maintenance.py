#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.5.9.2 - Inventory Maintenance Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.9: Inventory of Information and Assets
Assessment Domain 2 of 5: Inventory Maintenance & Update Procedures

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific inventory systems, update workflows, and integration
architecture.

Key customization areas:
1. Inventory system architecture (CMDB, spreadsheets, database, commercial tools)
2. Update workflows and triggers (procurement, decommissioning, change management)
3. Integration points (HR, procurement, asset management, monitoring tools)
4. SLA definitions and thresholds (what's acceptable for your operations)
5. Organization name, CISO details, contact information
6. File paths and naming conventions
7. Existing integration patterns and APIs

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
how well [Organization] maintains its asset inventory - the processes, systems,
and integrations that keep inventory data current and accurate over time.

**Purpose:**
Enables systematic assessment of inventory maintenance procedures, update
timeliness, system integrations, and quality control processes. Measures
operational effectiveness of keeping the inventory synchronized with actual
infrastructure changes, personnel moves, and procurement activities.

**Assessment Scope:**
- Inventory structure and access controls (who can view/edit what)
- Update triggers and workflows (how changes flow into inventory)
- Integration architecture (automated feeds from HR, procurement, monitoring)
- Quality control processes (validation, reconciliation, review cycles)
- Update timeliness and SLA compliance
- Documentation and procedures

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment methodology and maintenance best practices
2. Inventory Structure & Access - System architecture and access control model
3. Update Triggers & Workflows - How inventory stays synchronized with reality
4. Integration Architecture - Automated data feeds and system connections
5. Quality Control Processes - Validation, reconciliation, review procedures
6. Maintenance Metrics - Update timeliness, SLA compliance, effectiveness scores
7. Evidence Register - Process documentation and integration evidence

**Key Features:**
- Data validation with dropdown lists for system types and integration methods
- Conditional formatting for SLA compliance status (Green/Yellow/Red)
- Automated gap identification for missing integrations
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- SLA tracking and compliance measurement
- Integration health monitoring
- Update procedure documentation

**Integration:**
This assessment feeds into the A.5.9.5 Compliance Dashboard, which consolidates
data from all four assessment domains for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - os (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a59_2_inventory_maintenance.py

Output:
    File: ISMS-IMP-A.5.9.2_Inventory_Maintenance_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Complete maintenance assessment by documenting all update procedures
    2. Map integration architecture (HR, procurement, monitoring tools)
    3. Define and track SLAs for inventory updates
    4. Measure update timeliness and compliance
    5. Collect and link evidence in Evidence Register sheet
    6. Review with stakeholders (IT Ops, Asset Management, Procurement, HR)
    7. Export metrics CSV for dashboard consolidation (Sheet 6)
    8. Store assessment workbook per retention policy (7 years minimum)
    9. Update quarterly or after major process/system changes

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.9
Assessment Domain:    2 of 5
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 22.01.2026
Last Modified:        22.01.2026
Python Version:       3.8+
License:              [Organization License]

Related Documents:
    - ISMS-POL-A.5.9: Inventory of Information and Assets (Policy)
    - ISMS-IMP-A.5.9-1: Asset Discovery Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-2: Inventory Maintenance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-3: Quality & Compliance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-4: Owner Accountability Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 22.01.2026
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.9-2 specification
    - Inventory maintenance procedure documentation
    - Integration architecture mapping
    - SLA tracking and compliance measurement
    - Quality control process assessment
    - Evidence collection and audit trail support

[Future changes to be documented here]

--------------------------------------------------------------------------------
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

# CUSTOMIZE: Configuration
CONFIG = {
    'organization': '[Organization]',
    'ciso_name': '[CISO Name]',
    'security_contact': '[security@organization.com]',
    
    # Color scheme (consistent with A.8.24 pattern)
    'colors': {
        'header_bg': '003366',      # Dark blue
        'header_text': 'FFFFFF',     # White
        'section_bg': '4472C4',      # Medium blue
        'green_light': 'C6EFCE',     # Light green (pass)
        'green_dark': '006100',      # Dark green (pass text)
        'yellow_light': 'FFEB9C',    # Light yellow (warning)
        'yellow_dark': '9C5700',     # Dark yellow (warning text)
        'red_light': 'FFC7CE',       # Light red (fail)
        'red_dark': '9C0006',        # Dark red (fail text)
        'gray_light': 'D9D9D9',      # Light gray (locked cells)
    },
    
    # SLA targets (customize per organization)
    'sla_targets': {
        'New Asset Addition': 5,      # Business days
        'Asset Modification': 3,       # Business days
        'Asset Decommission': 2,       # Business days
        'Personnel Update (Joiner)': 1,    # Business days
        'Personnel Update (Leaver)': 1,    # Business days (CRITICAL)
        'Ownership Change': 3,         # Business days
    },
    
    # Update triggers
    'update_triggers': [
        'Procurement Approval',
        'Asset Received / Deployed',
        'Change Request Approved',
        'Decommission Request',
        'HR Onboarding',
        'HR Offboarding',
        'Ownership Transfer Request',
        'Discovery Scan Results',
        'Manual Entry',
        'Automated Integration',
        'Periodic Review',
        'Other (Specify)'
    ],
    
    # Integration types
    'integration_types': [
        'Real-time API',
        'Scheduled Batch (Daily)',
        'Scheduled Batch (Weekly)',
        'Webhook / Event-driven',
        'File Transfer (SFTP)',
        'Database Replication',
        'Manual Import',
        'No Integration (Manual)',
    ],
}


def main():
    """Main execution function"""
    print("="*80)
    print("ISMS Control A.5.9 - Inventory Maintenance Assessment Generator")
    print("="*80)
    print()
    print("Generating assessment workbook...")
    print()
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create sheets in order
    sheets = [
        "Instructions & Legend",
        "Inventory Structure & Access",
        "Update Triggers & Workflows",
        "Integration Architecture",
        "Quality Control Processes",
        "Maintenance Metrics",
        "Evidence Register",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
        print(f"  ✓ Created sheet: {sheet_name}")
    
    print()
    print("Populating sheets...")
    print()
    
    # Populate each sheet
    create_instructions_sheet(wb["Instructions & Legend"])
    print("  ✓ Instructions & Legend")
    
    create_inventory_structure_sheet(wb["Inventory Structure & Access"])
    print("  ✓ Inventory Structure & Access")
    
    create_update_triggers_sheet(wb["Update Triggers & Workflows"])
    print("  ✓ Update Triggers & Workflows")
    
    create_integration_architecture_sheet(wb["Integration Architecture"])
    print("  ✓ Integration Architecture")
    
    create_quality_control_sheet(wb["Quality Control Processes"])
    print("  ✓ Quality Control Processes")
    
    create_maintenance_metrics_sheet(wb["Maintenance Metrics"])
    print("  ✓ Maintenance Metrics")
    
    create_evidence_register_sheet(wb["Evidence Register"])
    print("  ✓ Evidence Register")
    
    # Save workbook
    filename = f"ISMS-IMP-A.5.9.2_Inventory_Maintenance_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    print()
    print("="*80)
    print(f"✅ SUCCESS: {filename}")
    print("="*80)
    print()
    print("Next Steps:")
    print("  1. Open workbook and review Instructions & Legend")
    print("  2. Document inventory system architecture and access controls")
    print("  3. Map all update triggers and workflows")
    print("  4. Document integration architecture (HR, procurement, monitoring)")
    print("  5. Assess quality control processes and procedures")
    print("  6. Review Maintenance Metrics for SLA compliance")
    print("  7. Export CSV from Sheet 6 for dashboard consolidation")
    print("  8. Obtain stakeholder review and approval")
    print()
    print(f"Output location: {os.path.abspath(filename)}")
    print()


def create_instructions_sheet(ws):
    """Create Instructions & Legend sheet"""
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "ISMS Control A.5.9 - Inventory Maintenance Assessment"
    ws['A1'].font = Font(size=16, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Subtitle
    ws.merge_cells('A2:H2')
    ws['A2'] = "Assessment Domain 2 of 5: Inventory Maintenance & Update Procedures"
    ws['A2'].font = Font(size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Purpose section
    ws['A4'] = "PURPOSE"
    ws['A4'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws['A4'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    ws.merge_cells('A5:H8')
    purpose_text = """This assessment evaluates HOW [Organization] keeps its asset inventory current and synchronized with actual infrastructure, personnel, and procurement changes.

Discovery finds assets that exist. Maintenance ensures the inventory STAYS accurate over time through systematic update procedures, integrations, and quality controls.

Four maintenance domains are assessed: Inventory Structure, Update Workflows, Integration Architecture, and Quality Control Processes."""
    ws['A5'] = purpose_text
    ws['A5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[5].height = 60
    
    # Assessment Methodology
    ws['A10'] = "ASSESSMENT METHODOLOGY"
    ws['A10'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws['A10'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    methodology_steps = [
        ("1. INVENTORY STRUCTURE", "Document system architecture, access controls, data model"),
        ("2. UPDATE TRIGGERS", "Identify all events that should trigger inventory updates (procurement, decom, HR changes)"),
        ("3. UPDATE WORKFLOWS", "Map procedures from trigger → inventory update → verification"),
        ("4. INTEGRATIONS", "Document automated data feeds (HR, procurement, monitoring tools, CMDB)"),
        ("5. QUALITY CONTROLS", "Assess validation rules, reconciliation processes, review cycles"),
        ("6. SLA COMPLIANCE", "Measure update timeliness vs. defined SLAs"),
    ]
    
    row = 11
    for step, description in methodology_steps:
        ws[f'A{row}'] = step
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = description
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1
    
    # SLA Targets
    ws[f'A{row+1}'] = "STANDARD SLA TARGETS"
    ws[f'A{row+1}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{row+1}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    row += 2
    ws[f'A{row}'] = "Update Trigger"
    ws[f'B{row}'] = "Target SLA (Business Days)"
    ws[f'C{row}'] = "Rationale"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = Font(bold=True)
    
    sla_data = [
        ("New Asset Addition", "5 days", "From procurement approval to inventory record"),
        ("Asset Modification", "3 days", "Configuration changes, relocations, upgrades"),
        ("Asset Decommission", "2 days", "Critical - must remove access quickly"),
        ("Personnel Joiner", "1 day", "HR onboarding triggers inventory update"),
        ("Personnel Leaver", "1 day", "CRITICAL - security risk if delayed"),
        ("Ownership Change", "3 days", "Ownership transfer approved to inventory updated"),
    ]
    
    row += 1
    for trigger, sla, rationale in sla_data:
        ws[f'A{row}'] = trigger
        ws[f'B{row}'] = sla
        ws[f'C{row}'] = rationale
        ws[f'C{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Traffic Light Legend
    ws[f'A{row+2}'] = "TRAFFIC LIGHT LEGEND"
    ws[f'A{row+2}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{row+2}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    row += 3
    # Green
    ws[f'A{row}'] = "✅ GREEN - Compliant"
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid')
    ws[f'B{row}'] = "SLA compliance ≥95% OR integration automated"
    
    # Yellow
    row += 1
    ws[f'A{row}'] = "⚠️ YELLOW - At Risk"
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    ws[f'B{row}'] = "SLA compliance 80-94% OR manual with documented procedure"
    
    # Red
    row += 1
    ws[f'A{row}'] = "❌ RED - Non-Compliant"
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid')
    ws[f'B{row}'] = "SLA compliance <80% OR no documented procedure"
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    
    # Protect sheet (read-only)
    ws.protection.sheet = True


def create_inventory_structure_sheet(ws):
    """Create Inventory Structure & Access sheet"""
    
    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "Inventory Structure & Access Controls"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Section 1: Inventory System Architecture
    ws['A3'] = "SECTION 1: INVENTORY SYSTEM ARCHITECTURE"
    ws['A3'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws['A3'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    ws.merge_cells('A3:L3')
    
    # Column headers
    headers = [
        ('A', 'System Component', 30),
        ('B', 'Technology/Platform', 25),
        ('C', 'Purpose', 40),
        ('D', 'Data Stored', 40),
        ('E', 'Owner', 20),
        ('F', 'Backup Frequency', 20),
        ('G', 'Availability', 15),
        ('H', 'Documentation', 30),
        ('I', 'Last Review', 15),
        ('J', 'Status', 15),
        ('K', 'Evidence ID', 15),
        ('L', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}4'] = header
        ws[f'{col}4'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}4'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws[f'{col}4'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[col].width = width
    
    # Example system components (user fills actual values)
    components = [
        ("Asset Inventory Database", "PostgreSQL / MySQL / SQL Server / Other", "Master inventory database"),
        ("CMDB (Configuration Management DB)", "ServiceNow / BMC / Jira / Other", "IT infrastructure CMDB"),
        ("Asset Management Tool", "Commercial Tool / Spreadsheet / Custom", "Asset tracking and management"),
        ("HR System Integration", "Workday / SAP SuccessFactors / Other", "Personnel asset integration"),
        ("Procurement System Integration", "SAP / Oracle / Coupa / Other", "Purchase order integration"),
    ]
    
    row = 5
    for component, tech, purpose in components:
        ws[f'A{row}'] = component
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = tech
        ws[f'C{row}'] = purpose
        
        # Unlock user input cells
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Add blank rows for additional components
    for i in range(5):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    # Section 2: Access Control Matrix
    row += 2
    ws[f'A{row}'] = "SECTION 2: ACCESS CONTROL MATRIX"
    ws[f'A{row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    ws.merge_cells(f'A{row}:L{row}')
    
    # Access control headers
    row += 1
    access_headers = [
        ('A', 'Role / User Group', 30),
        ('B', 'View Access', 15),
        ('C', 'Create Access', 15),
        ('D', 'Modify Access', 15),
        ('E', 'Delete Access', 15),
        ('F', 'Export Access', 15),
        ('G', 'Admin Access', 15),
        ('H', 'Justification', 40),
        ('I', 'Approved By', 20),
        ('J', 'Last Review', 15),
        ('K', 'Evidence ID', 15),
        ('L', 'Notes', 30),
    ]
    
    for col, header, width in access_headers:
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}{row}'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Example roles
    roles = [
        "Asset Owners",
        "IT Operations",
        "Security Team",
        "Auditors (Read-Only)",
        "Executive Management (Reports)",
        "Procurement Team",
        "HR Team",
    ]
    
    row += 1
    for role in roles:
        ws[f'A{row}'] = role
        ws[f'A{row}'].font = Font(bold=True)
        
        # Unlock for user input
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    # Data validations for Yes/No
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_yesno.add(f'B{row-7}:G{row-1}')
    ws.add_data_validation(dv_yesno)
    
    ws.protection.sheet = True


def create_update_triggers_sheet(ws):
    """Create Update Triggers & Workflows sheet"""
    
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "Update Triggers & Workflows"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25
    
    # Column headers
    headers = [
        ('A', 'Update Trigger Event', 35),
        ('B', 'Trigger Source', 25),
        ('C', 'Workflow Documented?', 20),
        ('D', 'Automation Level', 20),
        ('E', 'Target SLA (Days)', 15),
        ('F', 'Actual Avg Time (Days)', 18),
        ('G', 'SLA Compliance %', 16),
        ('H', 'Compliance Status', 18),
        ('I', 'Responsible Party', 25),
        ('J', 'Procedure Location', 35),
        ('K', 'Last Review', 15),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws[f'{col}3'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[col].width = width
    
    # Pre-populate trigger events
    trigger_events = [
        ("New Asset Procurement", "Procurement System", 5),
        ("Asset Deployment / Go-Live", "Change Management", 3),
        ("Asset Configuration Change", "Change Management", 3),
        ("Asset Relocation / Transfer", "Facilities / IT Ops", 3),
        ("Asset Decommission / Disposal", "Change Management", 2),
        ("Personnel Onboarding (Joiner)", "HR System", 1),
        ("Personnel Offboarding (Leaver)", "HR System", 1),
        ("Personnel Role Change", "HR System", 2),
        ("Ownership Transfer", "Asset Owner / Security", 3),
        ("Security Incident (Asset Affected)", "Security Team", 1),
        ("Discovery Scan Finds New Asset", "Network Monitoring", 5),
        ("Periodic Review Identifies Change", "Asset Owner", 7),
    ]
    
    row = 4
    for event, source, sla in trigger_events:
        ws[f'A{row}'] = event
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = source
        ws[f'E{row}'] = sla
        
        # Actual Avg Time (user enters from metrics)
        ws[f'F{row}'].protection = Protection(locked=False)
        
        # SLA Compliance % (formula)
        ws[f'G{row}'] = f'=IF(F{row}="","",IF(F{row}<=E{row},100,MAX(0,100-(F{row}-E{row})/E{row}*100)))'
        ws[f'G{row}'].number_format = '0"%"'
        
        # Compliance Status (formula)
        ws[f'H{row}'] = f'=IF(G{row}="","",IF(G{row}>=95,"✅ Compliant",IF(G{row}>=80,"⚠️ At Risk","❌ Non-Compliant")))'
        ws[f'H{row}'].alignment = Alignment(horizontal='center')
        
        # Unlock user input cells
        for col in ['C', 'D', 'F', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Data validations
    dv_yesno = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    dv_yesno.add(f'C4:C{row-1}')
    ws.add_data_validation(dv_yesno)
    
    automation_levels = ["Fully Automated", "Partially Automated", "Manual with Procedure", "Manual (Ad-hoc)", "Not Defined"]
    dv_automation = DataValidation(type="list", formula1=f'"{",".join(automation_levels)}"', allow_blank=True)
    dv_automation.add(f'D4:D{row-1}')
    ws.add_data_validation(dv_automation)
    
    # Summary section
    summary_row = row + 2
    ws[f'A{summary_row}'] = "UPDATE WORKFLOW SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Update Triggers"
    ws[f'B{summary_row}'] = f'=COUNTA(A4:A{row-1})'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Triggers with Documented Workflow"
    ws[f'B{summary_row}'] = f'=COUNTIF(C4:C{row-1},"Yes")'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Workflow Documentation %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-2}*100,0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Fully Automated Triggers"
    ws[f'B{summary_row}'] = f'=COUNTIF(D4:D{row-1},"Fully Automated")'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Automation Rate %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-4}*100,0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average SLA Compliance %"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(G4:G{row-1}),0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'G4:G{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['95'],
                   fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'G4:G{row-1}',
        CellIsRule(operator='between', formula=['80', '94'],
                   fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'G4:G{row-1}',
        CellIsRule(operator='lessThan', formula=['80'],
                   fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid'))
    )
    
    ws.protection.sheet = True


def create_integration_architecture_sheet(ws):
    """Create Integration Architecture sheet"""
    
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "Integration Architecture"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Source System', 30),
        ('B', 'Integration Type', 25),
        ('C', 'Data Flow', 30),
        ('D', 'Frequency', 20),
        ('E', 'Status', 15),
        ('F', 'Last Successful Run', 18),
        ('G', 'Success Rate %', 15),
        ('H', 'Health Status', 18),
        ('I', 'Owner', 20),
        ('J', 'Documentation', 35),
        ('K', 'Monitoring', 25),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Pre-populate integration examples
    integrations = [
        ("HR System (Personnel)", "Scheduled Batch (Daily)", "Employee data → Personnel Assets", "Daily 02:00"),
        ("Procurement System", "Webhook / Event-driven", "PO Approval → New Asset Record", "Real-time"),
        ("Network Monitoring (Nmap/Nessus)", "Scheduled Batch (Weekly)", "Scan Results → IT Infrastructure", "Weekly Sunday"),
        ("CMDB / ServiceNow", "Real-time API", "CI Updates → Inventory Sync", "Real-time"),
        ("Cloud Provider APIs (AWS/Azure)", "Scheduled Batch (Daily)", "Cloud Resources → IT Infrastructure", "Daily 03:00"),
        ("Active Directory", "Database Replication", "User/Group Data → Personnel/Access", "Hourly"),
        ("Asset Management Tool", "Real-time API", "Asset Changes → Inventory Sync", "Real-time"),
    ]
    
    row = 4
    for source, integration_type, data_flow, frequency in integrations:
        ws[f'A{row}'] = source
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = integration_type
        ws[f'C{row}'] = data_flow
        ws[f'D{row}'] = frequency
        
        # Success Rate % (user enters from monitoring)
        ws[f'G{row}'].protection = Protection(locked=False)
        ws[f'G{row}'].number_format = '0"%"'
        
        # Health Status (formula based on success rate)
        ws[f'H{row}'] = f'=IF(G{row}="","",IF(G{row}>=98,"✅ Healthy",IF(G{row}>=90,"⚠️ Degraded","❌ Unhealthy")))'
        ws[f'H{row}'].alignment = Alignment(horizontal='center')
        
        # Unlock user input cells
        for col in ['E', 'F', 'G', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Add blank rows
    for i in range(5):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    # Data validations
    dv_integration = DataValidation(type="list", formula1=f'"{",".join(CONFIG["integration_types"])}"', allow_blank=True)
    dv_integration.add(f'B4:B{row-1}')
    ws.add_data_validation(dv_integration)
    
    statuses = ["Active", "Inactive", "Planned", "Deprecated", "Failed"]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(statuses)}"', allow_blank=True)
    dv_status.add(f'E4:E{row-1}')
    ws.add_data_validation(dv_status)
    
    # Summary section
    summary_row = row + 2
    ws[f'A{summary_row}'] = "INTEGRATION HEALTH SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Integrations"
    ws[f'B{summary_row}'] = f'=COUNTA(A4:A{row-6})'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Active Integrations"
    ws[f'B{summary_row}'] = f'=COUNTIF(E4:E{row-6},"Active")'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Healthy Integrations"
    ws[f'B{summary_row}'] = f'=COUNTIF(H4:H{row-6},"✅ Healthy")'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Integration Health %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-2}*100,0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Success Rate"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(G4:G{row-6}),0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'G4:G{row-6}',
        CellIsRule(operator='greaterThanOrEqual', formula=['98'],
                   fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'G4:G{row-6}',
        CellIsRule(operator='between', formula=['90', '97'],
                   fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'G4:G{row-6}',
        CellIsRule(operator='lessThan', formula=['90'],
                   fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid'))
    )
    
    ws.protection.sheet = True


def create_quality_control_sheet(ws):
    """Create Quality Control Processes sheet"""
    
    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "Quality Control Processes"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Quality Control Process', 35),
        ('B', 'Frequency', 20),
        ('C', 'Documented?', 15),
        ('D', 'Last Performed', 18),
        ('E', 'Responsible Party', 25),
        ('F', 'Issues Found (Last Run)', 25),
        ('G', 'Issues Resolved', 20),
        ('H', 'Effectiveness %', 18),
        ('I', 'Status', 18),
        ('J', 'Evidence ID', 15),
        ('K', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Quality control processes
    qc_processes = [
        ("Data Validation Rules (Automated)", "Real-time", "Prevents invalid data entry"),
        ("Duplicate Detection / Deduplication", "Weekly", "Identifies and merges duplicate records"),
        ("Orphan Record Detection", "Weekly", "Finds assets without owners or parent records"),
        ("Consistency Checks (Cross-System)", "Monthly", "Compares inventory vs. HR, CMDB, procurement"),
        ("Completeness Checks (Mandatory Fields)", "Daily", "Identifies records missing required attributes"),
        ("Ownership Verification", "Quarterly", "Confirms asset owners are still valid and acknowledged"),
        ("Stale Record Review", "Monthly", "Identifies assets not reviewed in >90 days"),
        ("Access Control Audit", "Quarterly", "Reviews who has access to inventory systems"),
        ("Integration Reconciliation", "Weekly", "Verifies integrations running successfully"),
        ("Periodic Full Inventory Reconciliation", "Annually", "Complete audit of inventory vs. reality"),
    ]
    
    row = 4
    for process, frequency, description in qc_processes:
        ws[f'A{row}'] = process
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = frequency
        
        # Issues Found (user enters from last run)
        ws[f'F{row}'].protection = Protection(locked=False)
        
        # Issues Resolved
        ws[f'G{row}'].protection = Protection(locked=False)
        
        # Effectiveness % (formula: resolved / found)
        ws[f'H{row}'] = f'=IF(OR(F{row}="",G{row}=""),"",IF(F{row}=0,100,G{row}/F{row}*100))'
        ws[f'H{row}'].number_format = '0"%"'
        
        # Status
        ws[f'I{row}'] = f'=IF(H{row}="","",IF(H{row}>=90,"✅ Effective",IF(H{row}>=70,"⚠️ Needs Improvement","❌ Ineffective")))'
        ws[f'I{row}'].alignment = Alignment(horizontal='center')
        
        # Unlock user input cells
        for col in ['C', 'D', 'E', 'F', 'G', 'J', 'K']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Data validations
    dv_yesno = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    dv_yesno.add(f'C4:C{row-1}')
    ws.add_data_validation(dv_yesno)
    
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "QUALITY CONTROL SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total QC Processes"
    ws[f'B{summary_row}'] = f'=COUNTA(A4:A{row-1})'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Documented Processes"
    ws[f'B{summary_row}'] = f'=COUNTIF(C4:C{row-1},"Yes")'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Documentation Rate %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-2}*100,0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Effectiveness %"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(H4:H{row-1}),0)'
    ws[f'B{summary_row}'].number_format = '0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'H4:H{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['90'],
                   fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'H4:H{row-1}',
        CellIsRule(operator='between', formula=['70', '89'],
                   fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'H4:H{row-1}',
        CellIsRule(operator='lessThan', formula=['70'],
                   fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid'))
    )
    
    ws.protection.sheet = True


def create_maintenance_metrics_sheet(ws):
    """Create Maintenance Metrics sheet - consolidates from other sheets"""
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Maintenance Metrics & Summary"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Section 1: Overall Maintenance Effectiveness
    ws['A3'] = "OVERALL MAINTENANCE EFFECTIVENESS"
    ws['A3'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws['A3'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    ws.merge_cells('A3:H3')
    
    # Metrics table
    metrics = [
        ("Update Workflow SLA Compliance", "='Update Triggers & Workflows'!B40", "95%", "Update timeliness"),
        ("Integration Health", "='Integration Architecture'!B21", "98%", "Automated feed reliability"),
        ("Quality Control Effectiveness", "='Quality Control Processes'!B27", "90%", "Issue remediation rate"),
        ("Workflow Documentation Rate", "='Update Triggers & Workflows'!B36", "100%", "Procedure documentation"),
    ]
    
    row = 4
    for metric, formula, target, description in metrics:
        ws[f'A{row}'] = metric
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = '0"%"'
        ws[f'C{row}'] = target
        ws[f'D{row}'] = f'=B{row}-VALUE(LEFT(C{row},LEN(C{row})-1))/100'
        ws[f'D{row}'].number_format = '0.0"%"'
        ws[f'E{row}'] = f'=IF(B{row}>=VALUE(LEFT(C{row},LEN(C{row})-1))/100,"✅ Met",IF(B{row}>=VALUE(LEFT(C{row},LEN(C{row})-1))/100-0.1,"⚠️ At Risk","❌ Not Met"))'
        ws[f'E{row}'].alignment = Alignment(horizontal='center')
        ws[f'F{row}'] = description
        row += 1
    
    # Column headers for metrics table
    ws['A4'].value = "Metric"
    ws['B4'].value = "Actual"
    ws['C4'].value = "Target"
    ws['D4'].value = "Gap"
    ws['E4'].value = "Status"
    ws['F4'].value = "Description"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}4'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}4'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    
    row = 5
    for metric, formula, target, description in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = '0.0"%"'
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = target
        ws[f'D{row}'] = f'=B{row}-VALUE(LEFT(C{row},LEN(C{row})-1))/100'
        ws[f'D{row}'].number_format = '0.0"%"'
        ws[f'E{row}'] = f'=IF(B{row}>=VALUE(LEFT(C{row},LEN(C{row})-1))/100,"✅ Met",IF(B{row}>=VALUE(LEFT(C{row},LEN(C{row})-1))/100-0.1,"⚠️ At Risk","❌ Not Met"))'
        ws[f'E{row}'].alignment = Alignment(horizontal='center')
        ws[f'F{row}'] = description
        row += 1
    
    # Overall score
    overall_row = row + 2
    ws[f'A{overall_row}'] = "OVERALL MAINTENANCE SCORE"
    ws[f'A{overall_row}'].font = Font(size=12, bold=True)
    ws[f'B{overall_row}'] = f'=AVERAGE(B5:B8)'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    ws[f'B{overall_row}'].font = Font(bold=True, size=14)
    ws[f'B{overall_row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Target"
    ws[f'B{overall_row}'] = "95%"
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Gap"
    ws[f'B{overall_row}'] = f'=B{overall_row-2}-0.95'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    
    # CSV Export section
    csv_row = overall_row + 3
    ws[f'A{csv_row}'] = "CSV EXPORT FOR DASHBOARD (Copy rows below)"
    ws[f'A{csv_row}'].font = Font(size=11, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{csv_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    csv_row += 1
    ws[f'A{csv_row}'] = "Metric_Category"
    ws[f'B{csv_row}'] = "Score_%"
    ws[f'C{csv_row}'] = "Status"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{csv_row}'].font = Font(bold=True)
    
    csv_row += 1
    for i in range(4):
        ws[f'A{csv_row+i}'] = f'=A{5+i}'
        ws[f'B{csv_row+i}'] = f'=B{5+i}'
        ws[f'C{csv_row+i}'] = f'=E{5+i}'
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 40
    
    ws.protection.sheet = True


def create_evidence_register_sheet(ws):
    """Create Evidence Register sheet"""
    
    # Title
    ws.merge_cells('A1:N1')
    ws['A1'] = "Maintenance Evidence Register"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Evidence ID', 15),
        ('B', 'Maintenance Area', 30),
        ('C', 'Evidence Type', 30),
        ('D', 'Evidence Description', 50),
        ('E', 'Evidence Location', 40),
        ('F', 'Collection Date', 15),
        ('G', 'Collected By', 25),
        ('H', 'Validity Period', 20),
        ('I', 'Review Date', 15),
        ('J', 'Reviewed By', 25),
        ('K', 'Review Status', 20),
        ('L', 'Retention End Date', 18),
        ('M', 'Related Assessment', 25),
        ('N', 'Notes', 40),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Sample evidence
    sample_data = [
        ("MAINT-001", "Inventory Structure", "System Architecture Diagram", "Inventory database schema and access control model", "/evidence/inventory_architecture_20260122.pdf", "22.01.2026"),
        ("MAINT-002", "Update Workflows", "Procedure Document", "New asset onboarding workflow (procurement to inventory)", "/evidence/onboarding_procedure_v2.0.pdf", "22.01.2026"),
        ("MAINT-003", "Integration", "API Documentation", "HR system integration specification and data mapping", "/evidence/hr_integration_spec_20260122.pdf", "22.01.2026"),
        ("MAINT-004", "Quality Control", "Reconciliation Report", "Monthly inventory reconciliation vs. CMDB", "/evidence/reconciliation_report_202601.xlsx", "22.01.2026"),
    ]
    
    row = 4
    for evidence_id, area, evidence_type, description, location, date in sample_data:
        ws[f'A{row}'] = evidence_id
        ws[f'B{row}'] = area
        ws[f'C{row}'] = evidence_type
        ws[f'D{row}'] = description
        ws[f'E{row}'] = location
        ws[f'F{row}'] = date
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    # Add empty rows
    for i in range(20):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    # Data validations
    maintenance_areas = ["Inventory Structure", "Update Workflows", "Integration", "Quality Control", "All Areas"]
    dv_area = DataValidation(type="list", formula1=f'"{",".join(maintenance_areas)}"', allow_blank=True)
    dv_area.add(f'B4:B100')
    ws.add_data_validation(dv_area)
    
    review_statuses = ["Pending Review", "Reviewed - Valid", "Reviewed - Update Needed", "Reviewed - Invalid"]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(review_statuses)}"', allow_blank=True)
    dv_status.add(f'K4:K100')
    ws.add_data_validation(dv_status)
    
    ws.protection.sheet = True


# Execute main function
if __name__ == "__main__":
    main()
