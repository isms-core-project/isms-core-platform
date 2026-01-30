#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.5.9.1 - Asset Discovery Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.9: Inventory of Information and Assets
Assessment Domain 1 of 5: Asset Discovery & Completeness Verification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific asset landscape, discovery tools, and assessment
requirements.

Key customization areas:
1. Asset categories and types (match your business context and infrastructure)
2. Discovery methods and tools (network scanning, HR systems, procurement, etc.)
3. Completeness thresholds (what percentage is acceptable for your risk profile)
4. Organization name, CISO details, contact information
5. File paths and naming conventions
6. Integration with existing asset management/CMDB systems

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
asset discovery completeness across all five asset categories defined in ISO
27001:2022 Control A.5.9.

**Purpose:**
Enables systematic assessment of whether [Organization] has identified and
documented all assets that store, process, transmit, or protect information.
Measures discovery completeness against expected asset populations to identify
gaps and guide remediation efforts.

**Assessment Scope:**
- Information Assets (databases, documents, records, intellectual property)
- IT Infrastructure (servers, network devices, storage, endpoints)
- Applications (SaaS, on-premise, custom, COTS)
- Physical Assets (facilities, equipment, media, backup tapes)
- Personnel Assets (roles, competencies, critical knowledge holders)

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment methodology and completion guidance
2. Information Assets Discovery - Discovery completeness for information assets
3. IT Infrastructure Discovery - Discovery completeness for IT assets
4. Applications Discovery - Discovery completeness for applications
5. Physical Assets Discovery - Discovery completeness for physical assets
6. Personnel Assets Discovery - Discovery completeness for personnel assets
7. Discovery Metrics & Summary - Aggregated completeness scores and gaps
8. Evidence Register - Discovery evidence tracking and documentation

**Key Features:**
- Data validation with dropdown lists for asset categories
- Conditional formatting for completeness status (Green/Yellow/Red)
- Automated gap identification and prioritization
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Discovery method tracking (network scan, manual, procurement, HR)
- Completeness percentage calculations by category
- Gap analysis with remediation recommendations

**Integration:**
This assessment feeds into the A.5.9.5 Compliance Dashboard, which consolidates
data from all four assessment domains for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - os (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a59_1_asset_discovery.py

Output:
    File: ISMS-IMP-A.5.9.1_Asset_Discovery_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Complete discovery assessment by populating all asset categories
    2. Document discovery methods used (network scans, HR exports, etc.)
    3. Calculate completeness percentages vs. expected populations
    4. Identify gaps and prioritize discovery efforts
    5. Collect and link evidence in Evidence Register sheet
    6. Review with stakeholders (IT, HR, Security, Asset Owners)
    7. Export metrics CSV for dashboard consolidation (Sheet 7)
    8. Store assessment workbook per retention policy (7 years minimum)
    9. Update quarterly or after major infrastructure/organizational changes

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.9
Assessment Domain:    1 of 5
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 22.01.2026
Last Modified:        22.01.2026
Python Version:       3.8+
License:              [Organization License]

Related Documents:
    - ISMS-POL-A.5.9: Inventory of Information and Assets (Policy)
    - ISMS-IMP-A.5.9-1: Asset Discovery Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-2: Inventory Maintenance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-3: Quality & Compliance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-4: Owner Accountability Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 22.01.2026
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.9-1 specification
    - Discovery completeness tracking across 5 asset categories
    - Gap identification and remediation planning
    - Evidence collection and audit trail support
    - CSV export for dashboard consolidation

[Future changes to be documented here]

--------------------------------------------------------------------------------
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

# CUSTOMIZE: Configuration
CONFIG = {
    'organization': '[Organization]',
    'ciso_name': '[CISO Name]',
    'security_contact': '[security@organization.com]',
    
    # Color scheme (consistent with A.8.24 pattern)
    'colors': {
        'header_bg': '003366',      # Dark blue
        'header_text': 'FFFFFF',     # White
        'section_bg': '4472C4',      # Medium blue
        'green_light': 'C6EFCE',     # Light green (pass)
        'green_dark': '006100',      # Dark green (pass text)
        'yellow_light': 'FFEB9C',    # Light yellow (warning)
        'yellow_dark': '9C5700',     # Dark yellow (warning text)
        'red_light': 'FFC7CE',       # Light red (fail)
        'red_dark': '9C0006',        # Dark red (fail text)
        'gray_light': 'D9D9D9',      # Light gray (locked cells)
    },
    
    # Completeness targets (customize per organization risk appetite)
    'targets': {
        'Information Assets': 95,
        'IT Infrastructure': 98,
        'Applications': 90,
        'Physical Assets': 90,
        'Personnel Assets': 100,  # Critical - must know all key personnel
    },
    
    # Discovery methods
    'discovery_methods': [
        'Network Scan (Automated)',
        'CMDB/Asset Management System',
        'Procurement Records',
        'HR System Export',
        'Manual Survey',
        'License Audit',
        'Physical Inspection',
        'Department Interviews',
        'Document Repository Scan',
        'Cloud Provider API',
        'Other (Specify)'
    ],
}


def main():
    """Main execution function"""
    print("="*80)
    print("ISMS Control A.5.9 - Asset Discovery Assessment Generator")
    print("="*80)
    print()
    print("Generating assessment workbook...")
    print()
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create sheets in order
    sheets = [
        "Instructions & Legend",
        "Information Assets Discovery",
        "IT Infrastructure Discovery",
        "Applications Discovery",
        "Physical Assets Discovery",
        "Personnel Assets Discovery",
        "Discovery Metrics & Summary",
        "Evidence Register",
        "Summary Dashboard",
        "Approval & Sign-Off",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
        print(f"  ✓ Created sheet: {sheet_name}")
    
    print()
    print("Populating sheets...")
    print()
    
    # Populate each sheet
    create_instructions_sheet(wb["Instructions & Legend"])
    print("  ✓ Instructions & Legend")
    
    create_information_assets_sheet(wb["Information Assets Discovery"])
    print("  ✓ Information Assets Discovery")
    
    create_it_infrastructure_sheet(wb["IT Infrastructure Discovery"])
    print("  ✓ IT Infrastructure Discovery")
    
    create_applications_sheet(wb["Applications Discovery"])
    print("  ✓ Applications Discovery")
    
    create_physical_assets_sheet(wb["Physical Assets Discovery"])
    print("  ✓ Physical Assets Discovery")
    
    create_personnel_assets_sheet(wb["Personnel Assets Discovery"])
    print("  ✓ Personnel Assets Discovery")
    
    create_discovery_metrics_sheet(wb["Discovery Metrics & Summary"])
    print("  ✓ Discovery Metrics & Summary")
    
    create_evidence_register_sheet(wb["Evidence Register"])
    print("  ✓ Evidence Register")
    
    create_summary_dashboard_sheet(wb["Summary Dashboard"])
    print("  ✓ Summary Dashboard")
    
    create_approval_signoff_sheet(wb["Approval & Sign-Off"])
    print("  ✓ Approval & Sign-Off")
    
    # Save workbook
    filename = f"ISMS-IMP-A.5.9.1_Asset_Discovery_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    print()
    print("="*80)
    print(f"✅ SUCCESS: {filename}")
    print("="*80)
    print()
    print("Next Steps:")
    print("  1. Open workbook and review Instructions & Legend")
    print("  2. Complete discovery assessment for each asset category")
    print("  3. Document discovery methods and evidence")
    print("  4. Review Discovery Metrics & Summary for completeness scores")
    print("  5. Export CSV from Sheet 7 for dashboard consolidation")
    print("  6. Obtain stakeholder review and approval")
    print()
    print(f"Output location: {os.path.abspath(filename)}")
    print()


def create_instructions_sheet(ws):
    """Create Instructions & Legend sheet"""
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "ISMS Control A.5.9 - Asset Discovery Assessment"
    ws['A1'].font = Font(size=16, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Subtitle
    ws.merge_cells('A2:H2')
    ws['A2'] = "Assessment Domain 1 of 5: Asset Discovery & Completeness Verification"
    ws['A2'].font = Font(size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Purpose section
    ws['A4'] = "PURPOSE"
    ws['A4'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws['A4'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    ws.merge_cells('A5:H8')
    purpose_text = """This assessment evaluates whether [Organization] has successfully identified and documented all assets that store, process, transmit, or protect information assets.

The goal is to measure discovery COMPLETENESS - do we know what exists? This is the foundation for all other inventory activities (maintenance, quality, accountability).

Five asset categories are assessed: Information Assets, IT Infrastructure, Applications, Physical Assets, and Personnel Assets."""
    ws['A5'] = purpose_text
    ws['A5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[5].height = 60
    
    # Assessment Methodology
    ws['A10'] = "ASSESSMENT METHODOLOGY"
    ws['A10'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws['A10'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    methodology_steps = [
        ("1. EXPECTED POPULATION", "Estimate total expected assets in each category (from business context, org size, infrastructure)"),
        ("2. DISCOVERED ASSETS", "Count assets actually identified and documented via discovery methods"),
        ("3. COMPLETENESS %", "Calculate: (Discovered / Expected) × 100%"),
        ("4. GAP ANALYSIS", "Identify categories <95% complete, prioritize additional discovery"),
        ("5. DISCOVERY METHODS", "Document how assets were found (network scan, HR export, manual, etc.)"),
    ]
    
    row = 11
    for step, description in methodology_steps:
        ws[f'A{row}'] = step
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = description
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1
    
    # Completeness Targets
    ws[f'A{row+1}'] = "COMPLETENESS TARGETS"
    ws[f'A{row+1}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{row+1}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    row += 2
    ws[f'A{row}'] = "Asset Category"
    ws[f'B{row}'] = "Target Completeness"
    ws[f'C{row}'] = "Rationale"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = Font(bold=True)
    
    targets_data = [
        ("Information Assets", "95%", "High criticality - must know sensitive data locations"),
        ("IT Infrastructure", "98%", "Critical for security controls - servers, network devices"),
        ("Applications", "90%", "Acceptable - shadow IT discovery ongoing"),
        ("Physical Assets", "90%", "Facilities and equipment - manual discovery acceptable"),
        ("Personnel Assets", "100%", "Critical - must know all key personnel and competencies"),
    ]
    
    row += 1
    for category, target, rationale in targets_data:
        ws[f'A{row}'] = category
        ws[f'B{row}'] = target
        ws[f'C{row}'] = rationale
        ws[f'C{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Traffic Light Legend
    ws[f'A{row+2}'] = "TRAFFIC LIGHT LEGEND"
    ws[f'A{row+2}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{row+2}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    row += 3
    # Green
    ws[f'A{row}'] = "✅ GREEN - Compliant"
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid')
    ws[f'B{row}'] = "Completeness ≥ Target (e.g., ≥95%)"
    
    # Yellow
    row += 1
    ws[f'A{row}'] = "⚠️ YELLOW - At Risk"
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    ws[f'B{row}'] = "Completeness within 10% of target (e.g., 85-94%)"
    
    # Red
    row += 1
    ws[f'A{row}'] = "❌ RED - Non-Compliant"
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid')
    ws[f'B{row}'] = "Completeness >10% below target (e.g., <85%)"
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    
    # Protect sheet (read-only, no password)
    ws.protection.sheet = True  # No password needed for instructions


def create_information_assets_sheet(ws):
    """Create Information Assets Discovery sheet"""
    
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "Information Assets Discovery Assessment"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Column headers (row 3)
    headers = [
        ('A', 'Asset Subcategory', 25),
        ('B', 'Discovery Method', 30),
        ('C', 'Expected Count', 15),
        ('D', 'Discovered Count', 15),
        ('E', 'Completeness %', 15),
        ('F', 'Compliance Status', 18),
        ('G', 'Gaps Identified', 40),
        ('H', 'Discovery Evidence', 30),
        ('I', 'Next Discovery Actions', 40),
        ('J', 'Responsible Party', 25),
        ('K', 'Target Date', 15),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws[f'{col}3'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[col].width = width
    
    ws.row_dimensions[3].height = 30
    
    # Information Asset Subcategories (pre-populated examples - customize)
    subcategories = [
        "Databases (Production)",
        "Databases (Development/Test)",
        "Data Warehouses / Analytics",
        "File Shares (Departmental)",
        "SharePoint / Document Management",
        "Business Records / Archives",
        "Email Archives",
        "Intellectual Property (Patents, Trade Secrets)",
        "Source Code Repositories",
        "Customer Data Collections",
        "Financial Records",
        "HR Records (Personnel Files)",
        "Contracts and Legal Documents",
        "Technical Documentation",
        "Marketing / Sales Materials",
    ]
    
    row = 4
    for subcategory in subcategories:
        ws[f'A{row}'] = subcategory
        
        # Expected Count (user enters)
        ws[f'C{row}'].protection = Protection(locked=False)
        
        # Discovered Count (user enters)
        ws[f'D{row}'].protection = Protection(locked=False)
        
        # Completeness % (formula)
        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        
        # Compliance Status (formula - based on 95% target)
        ws[f'F{row}'] = f'=IF(E{row}>=95,"✅ Compliant",IF(E{row}>=85,"⚠️ At Risk","❌ Non-Compliant"))'
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        
        # Unlocked user input cells
        for col in ['B', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Discovery Method dropdown (column B)
    dv_method = DataValidation(
        type="list",
        formula1=f'"{",".join(CONFIG["discovery_methods"])}"',
        allow_blank=True
    )
    dv_method.add(f'B4:B{row-1}')
    ws.add_data_validation(dv_method)
    
    # Summary section
    summary_row = row + 2
    ws[f'A{summary_row}'] = "INFORMATION ASSETS SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Expected"
    ws[f'B{summary_row}'] = f'=SUM(C4:C{row-1})'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Discovered"
    ws[f'B{summary_row}'] = f'=SUM(D4:D{row-1})'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Overall Completeness %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-2}*100,0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "95%"
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap vs. Target"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-95'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting for Completeness %
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['95'],
                   fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='between', formula=['85', '94'],
                   fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='lessThan', formula=['85'],
                   fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid'))
    )
    
    # Protect sheet (formulas locked, inputs unlocked)


def create_it_infrastructure_sheet(ws):
    """Create IT Infrastructure Discovery sheet"""
    
    # Same structure as Information Assets but with IT-specific subcategories
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "IT Infrastructure Discovery Assessment"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Column headers
    headers = [
        ('A', 'Asset Subcategory', 25),
        ('B', 'Discovery Method', 30),
        ('C', 'Expected Count', 15),
        ('D', 'Discovered Count', 15),
        ('E', 'Completeness %', 15),
        ('F', 'Compliance Status', 18),
        ('G', 'Gaps Identified', 40),
        ('H', 'Discovery Evidence', 30),
        ('I', 'Next Discovery Actions', 40),
        ('J', 'Responsible Party', 25),
        ('K', 'Target Date', 15),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws[f'{col}3'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[col].width = width
    
    # IT Infrastructure Subcategories
    subcategories = [
        "Physical Servers (Production)",
        "Physical Servers (Non-Production)",
        "Virtual Servers / VMs",
        "Network Routers",
        "Network Switches",
        "Firewalls",
        "Load Balancers",
        "Storage Arrays / NAS / SAN",
        "Backup Devices / Tape Libraries",
        "Workstations / Desktops",
        "Laptops / Mobile Devices",
        "Tablets / iPads",
        "Smartphones (Corporate)",
        "Printers / Multifunction Devices",
        "IoT Devices / Sensors",
        "Network Attached Devices",
    ]
    
    row = 4
    for subcategory in subcategories:
        ws[f'A{row}'] = subcategory
        ws[f'C{row}'].protection = Protection(locked=False)
        ws[f'D{row}'].protection = Protection(locked=False)
        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        ws[f'F{row}'] = f'=IF(E{row}>=98,"✅ Compliant",IF(E{row}>=88,"⚠️ At Risk","❌ Non-Compliant"))'
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        
        for col in ['B', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    # Discovery Method dropdown
    dv_method = DataValidation(
        type="list",
        formula1=f'"{",".join(CONFIG["discovery_methods"])}"',
        allow_blank=True
    )
    dv_method.add(f'B4:B{row-1}')
    ws.add_data_validation(dv_method)
    
    # Summary section (98% target for IT Infrastructure)
    summary_row = row + 2
    ws[f'A{summary_row}'] = "IT INFRASTRUCTURE SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Expected"
    ws[f'B{summary_row}'] = f'=SUM(C4:C{row-1})'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Discovered"
    ws[f'B{summary_row}'] = f'=SUM(D4:D{row-1})'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Overall Completeness %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-2}*100,0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "98%"
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap vs. Target"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-98'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting (98% threshold)
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['98'],
                   fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='between', formula=['88', '97'],
                   fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='lessThan', formula=['88'],
                   fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid'))
    )


def create_applications_sheet(ws):
    """Create Applications Discovery sheet"""
    
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "Applications Discovery Assessment"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Column headers
    headers = [
        ('A', 'Asset Subcategory', 25),
        ('B', 'Discovery Method', 30),
        ('C', 'Expected Count', 15),
        ('D', 'Discovered Count', 15),
        ('E', 'Completeness %', 15),
        ('F', 'Compliance Status', 18),
        ('G', 'Gaps Identified', 40),
        ('H', 'Discovery Evidence', 30),
        ('I', 'Next Discovery Actions', 40),
        ('J', 'Responsible Party', 25),
        ('K', 'Target Date', 15),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Application Subcategories
    subcategories = [
        "SaaS Applications (Licensed)",
        "Cloud-Hosted Applications (Custom)",
        "On-Premise Applications (COTS)",
        "On-Premise Applications (Custom/In-House)",
        "Database Management Systems",
        "Middleware / Integration Platforms",
        "Development Tools / IDEs",
        "Security Tools (SIEM, EDR, Firewall)",
        "Monitoring / Observability Tools",
        "Backup / Recovery Software",
        "Collaboration Tools (Slack, Teams, etc.)",
        "Productivity Suites (Office 365, G Suite)",
        "ERP / CRM Systems",
        "HR Management Systems",
        "Shadow IT (Unauthorized SaaS)",
    ]
    
    row = 4
    for subcategory in subcategories:
        ws[f'A{row}'] = subcategory
        ws[f'C{row}'].protection = Protection(locked=False)
        ws[f'D{row}'].protection = Protection(locked=False)
        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        ws[f'F{row}'] = f'=IF(E{row}>=90,"✅ Compliant",IF(E{row}>=80,"⚠️ At Risk","❌ Non-Compliant"))'
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        
        for col in ['B', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    dv_method = DataValidation(type="list", formula1=f'"{",".join(CONFIG["discovery_methods"])}"', allow_blank=True)
    dv_method.add(f'B4:B{row-1}')
    ws.add_data_validation(dv_method)
    
    # Summary (90% target)
    summary_row = row + 2
    ws[f'A{summary_row}'] = "APPLICATIONS SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    ws[f'A{summary_row+1}'] = "Total Expected"
    ws[f'B{summary_row+1}'] = f'=SUM(C4:C{row-1})'
    ws[f'B{summary_row+1}'].font = Font(bold=True)
    
    ws[f'A{summary_row+2}'] = "Total Discovered"
    ws[f'B{summary_row+2}'] = f'=SUM(D4:D{row-1})'
    ws[f'B{summary_row+2}'].font = Font(bold=True)
    
    ws[f'A{summary_row+3}'] = "Overall Completeness %"
    ws[f'B{summary_row+3}'] = f'=IFERROR(B{summary_row+2}/B{summary_row+1}*100,0)'
    ws[f'B{summary_row+3}'].number_format = '0.0"%"'
    ws[f'B{summary_row+3}'].font = Font(bold=True, size=12)
    
    ws[f'A{summary_row+4}'] = "Target"
    ws[f'B{summary_row+4}'] = "90%"
    ws[f'B{summary_row+4}'].font = Font(bold=True)
    
    ws[f'A{summary_row+5}'] = "Gap vs. Target"
    ws[f'B{summary_row+5}'] = f'=B{summary_row+3}-90'
    ws[f'B{summary_row+5}'].number_format = '0.0"%"'
    
    # Conditional formatting
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='greaterThanOrEqual', formula=['90'], fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='between', formula=['80', '89'], fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='lessThan', formula=['80'], fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid')))


def create_physical_assets_sheet(ws):
    """Create Physical Assets Discovery sheet"""
    
    # Similar structure with physical asset categories
    ws.merge_cells('A1:M1')
    ws['A1'] = "Physical Assets Discovery Assessment"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Asset Subcategory', 25),
        ('B', 'Discovery Method', 30),
        ('C', 'Expected Count', 15),
        ('D', 'Discovered Count', 15),
        ('E', 'Completeness %', 15),
        ('F', 'Compliance Status', 18),
        ('G', 'Gaps Identified', 40),
        ('H', 'Discovery Evidence', 30),
        ('I', 'Next Discovery Actions', 40),
        ('J', 'Responsible Party', 25),
        ('K', 'Target Date', 15),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    subcategories = [
        "Data Centers / Server Rooms",
        "Office Facilities",
        "Storage Facilities (Offsite)",
        "Backup Tape Libraries (Offsite)",
        "Removable Media (USB Drives, External HDDs)",
        "Paper Records / File Cabinets",
        "Safes / Secure Storage",
        "Badge Readers / Access Control Systems",
        "CCTV / Surveillance Systems",
        "Environmental Controls (HVAC, Fire Suppression)",
        "Power Infrastructure (UPS, Generators)",
        "Network Cabling Infrastructure",
    ]
    
    row = 4
    for subcategory in subcategories:
        ws[f'A{row}'] = subcategory
        ws[f'C{row}'].protection = Protection(locked=False)
        ws[f'D{row}'].protection = Protection(locked=False)
        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        ws[f'F{row}'] = f'=IF(E{row}>=90,"✅ Compliant",IF(E{row}>=80,"⚠️ At Risk","❌ Non-Compliant"))'
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        
        for col in ['B', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    dv_method = DataValidation(type="list", formula1=f'"{",".join(CONFIG["discovery_methods"])}"', allow_blank=True)
    dv_method.add(f'B4:B{row-1}')
    ws.add_data_validation(dv_method)
    
    # Summary (90% target)
    summary_row = row + 2
    ws[f'A{summary_row}'] = "PHYSICAL ASSETS SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    ws[f'A{summary_row+1}'] = "Total Expected"
    ws[f'B{summary_row+1}'] = f'=SUM(C4:C{row-1})'
    ws[f'A{summary_row+2}'] = "Total Discovered"
    ws[f'B{summary_row+2}'] = f'=SUM(D4:D{row-1})'
    ws[f'A{summary_row+3}'] = "Overall Completeness %"
    ws[f'B{summary_row+3}'] = f'=IFERROR(B{summary_row+2}/B{summary_row+1}*100,0)'
    ws[f'B{summary_row+3}'].number_format = '0.0"%"'
    ws[f'B{summary_row+3}'].font = Font(bold=True, size=12)
    ws[f'A{summary_row+4}'] = "Target"
    ws[f'B{summary_row+4}'] = "90%"
    ws[f'A{summary_row+5}'] = "Gap vs. Target"
    ws[f'B{summary_row+5}'] = f'=B{summary_row+3}-90'
    ws[f'B{summary_row+5}'].number_format = '0.0"%"'
    
    # Conditional formatting
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='greaterThanOrEqual', formula=['90'], fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='between', formula=['80', '89'], fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='lessThan', formula=['80'], fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid')))


def create_personnel_assets_sheet(ws):
    """Create Personnel Assets Discovery sheet"""
    
    ws.merge_cells('A1:M1')
    ws['A1'] = "Personnel Assets Discovery Assessment"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Asset Subcategory', 25),
        ('B', 'Discovery Method', 30),
        ('C', 'Expected Count', 15),
        ('D', 'Discovered Count', 15),
        ('E', 'Completeness %', 15),
        ('F', 'Compliance Status', 18),
        ('G', 'Gaps Identified', 40),
        ('H', 'Discovery Evidence', 30),
        ('I', 'Next Discovery Actions', 40),
        ('J', 'Responsible Party', 25),
        ('K', 'Target Date', 15),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    subcategories = [
        "Executive Management (C-Suite)",
        "Department Heads / Managers",
        "Information Security Team",
        "IT Operations Team",
        "Developers / Engineers",
        "Database Administrators",
        "System Administrators",
        "Network Engineers",
        "Security Analysts / SOC",
        "Compliance Officers",
        "Data Protection Officers",
        "Key Subject Matter Experts",
        "Critical Project Personnel",
        "Contractors (Long-term)",
        "External Service Providers (Key Personnel)",
    ]
    
    row = 4
    for subcategory in subcategories:
        ws[f'A{row}'] = subcategory
        ws[f'C{row}'].protection = Protection(locked=False)
        ws[f'D{row}'].protection = Protection(locked=False)
        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        # 100% target for personnel - CRITICAL
        ws[f'F{row}'] = f'=IF(E{row}=100,"✅ Compliant",IF(E{row}>=90,"⚠️ At Risk","❌ Non-Compliant"))'
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        
        for col in ['B', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    dv_method = DataValidation(type="list", formula1=f'"{",".join(CONFIG["discovery_methods"])}"', allow_blank=True)
    dv_method.add(f'B4:B{row-1}')
    ws.add_data_validation(dv_method)
    
    # Summary (100% target - critical!)
    summary_row = row + 2
    ws[f'A{summary_row}'] = "PERSONNEL ASSETS SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    ws[f'A{summary_row+1}'] = "Total Expected"
    ws[f'B{summary_row+1}'] = f'=SUM(C4:C{row-1})'
    ws[f'A{summary_row+2}'] = "Total Discovered"
    ws[f'B{summary_row+2}'] = f'=SUM(D4:D{row-1})'
    ws[f'A{summary_row+3}'] = "Overall Completeness %"
    ws[f'B{summary_row+3}'] = f'=IFERROR(B{summary_row+2}/B{summary_row+1}*100,0)'
    ws[f'B{summary_row+3}'].number_format = '0.0"%"'
    ws[f'B{summary_row+3}'].font = Font(bold=True, size=12)
    ws[f'A{summary_row+4}'] = "Target"
    ws[f'B{summary_row+4}'] = "100%"
    ws[f'B{summary_row+4}'].font = Font(bold=True, color=CONFIG['colors']['red_dark'])
    ws[f'A{summary_row+5}'] = "Gap vs. Target"
    ws[f'B{summary_row+5}'] = f'=B{summary_row+3}-100'
    ws[f'B{summary_row+5}'].number_format = '0.0"%"'
    
    # Conditional formatting (stricter - 100% required)
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='equal', formula=['100'], fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='between', formula=['90', '99'], fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')))
    ws.conditional_formatting.add(f'E4:E{row-1}', CellIsRule(operator='lessThan', formula=['90'], fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid')))


def create_discovery_metrics_sheet(ws):
    """Create Discovery Metrics & Summary sheet - consolidates all 5 categories"""
    
    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "Discovery Metrics & Summary - All Asset Categories"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25
    
    # Column headers
    headers = [
        ('A', 'Asset Category', 30),
        ('B', 'Expected Count', 15),
        ('C', 'Discovered Count', 15),
        ('D', 'Completeness %', 15),
        ('E', 'Target %', 12),
        ('F', 'Gap vs. Target', 15),
        ('G', 'Compliance Status', 20),
        ('H', 'Priority', 15),
        ('I', 'Key Gaps', 50),
        ('J', 'Next Actions', 50),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Consolidate from other sheets
    categories = [
        ("Information Assets", "'Information Assets Discovery'", 95),
        ("IT Infrastructure", "'IT Infrastructure Discovery'", 98),
        ("Applications", "'Applications Discovery'", 90),
        ("Physical Assets", "'Physical Assets Discovery'", 90),
        ("Personnel Assets", "'Personnel Assets Discovery'", 100),
    ]
    
    row = 4
    for category_name, sheet_ref, target in categories:
        ws[f'A{row}'] = category_name
        ws[f'A{row}'].font = Font(bold=True)
        
        # Find summary row in source sheet (typically row with "Total Expected")
        # For simplicity, hardcode summary row positions (adjust based on actual sheets)
        # Information Assets: summary at row 22, IT: row 23, Apps: row 21, etc.
        # Using approximate positions - would need exact calculation
        
        # Expected Count (link to source sheet summary)
        ws[f'B{row}'] = f'={sheet_ref}!B22'  # Adjust row as needed
        
        # Discovered Count
        ws[f'C{row}'] = f'={sheet_ref}!B23'  # Adjust row as needed
        
        # Completeness %
        ws[f'D{row}'] = f'=IFERROR(C{row}/B{row}*100,0)'
        ws[f'D{row}'].number_format = '0.0"%"'
        ws[f'D{row}'].font = Font(bold=True)
        
        # Target
        ws[f'E{row}'] = f'{target}%'
        
        # Gap vs. Target
        ws[f'F{row}'] = f'=D{row}-{target}'
        ws[f'F{row}'].number_format = '0.0"%"'
        
        # Compliance Status (formula)
        ws[f'G{row}'] = f'=IF(D{row}>={target},"✅ Compliant",IF(D{row}>={target}-10,"⚠️ At Risk","❌ Non-Compliant"))'
        ws[f'G{row}'].alignment = Alignment(horizontal='center')
        
        # Priority (based on gap)
        ws[f'H{row}'] = f'=IF(F{row}<-10,"🔴 Critical",IF(F{row}<0,"🟡 High","🟢 Low"))'
        ws[f'H{row}'].alignment = Alignment(horizontal='center')
        
        # Key Gaps and Next Actions (user enters)
        ws[f'I{row}'].protection = Protection(locked=False)
        ws[f'J{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Overall Discovery Summary
    overall_row = row + 2
    ws[f'A{overall_row}'] = "OVERALL DISCOVERY SUMMARY"
    ws[f'A{overall_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{overall_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    ws.merge_cells(f'A{overall_row}:J{overall_row}')
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Total Assets Expected (All Categories)"
    ws[f'B{overall_row}'] = f'=SUM(B4:B8)'
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Total Assets Discovered"
    ws[f'B{overall_row}'] = f'=SUM(C4:C8)'
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Discovery Completeness %"
    ws[f'B{overall_row}'] = f'=IFERROR(B{overall_row-1}/B{overall_row-2}*100,0)'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    ws[f'B{overall_row}'].font = Font(bold=True, size=14)
    ws[f'B{overall_row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Target"
    ws[f'B{overall_row}'] = "95%"
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Gap vs. Overall Target"
    ws[f'B{overall_row}'] = f'=B{overall_row-2}-95'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    # CSV Export section for dashboard
    csv_row = overall_row + 3
    ws[f'A{csv_row}'] = "CSV EXPORT FOR DASHBOARD (Copy rows below)"
    ws[f'A{csv_row}'].font = Font(size=11, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{csv_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    csv_row += 1
    ws[f'A{csv_row}'] = "Discovery_Category"
    ws[f'B{csv_row}'] = "Completeness_%"
    ws[f'C{csv_row}'] = "Compliance_Status"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{csv_row}'].font = Font(bold=True)
    
    csv_row += 1
    # Link to data rows
    for i in range(5):
        ws[f'A{csv_row+i}'] = f'=A{4+i}'
        ws[f'B{csv_row+i}'] = f'=D{4+i}'
        ws[f'C{csv_row+i}'] = f'=G{4+i}'
    
    # Conditional formatting for Completeness %
    ws.conditional_formatting.add(
        'D4:D8',
        CellIsRule(operator='greaterThanOrEqual', formula=['95'],
                   fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        'D4:D8',
        CellIsRule(operator='between', formula=['85', '94'],
                   fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        'D4:D8',
        CellIsRule(operator='lessThan', formula=['85'],
                   fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid'))
    )


def create_evidence_register_sheet(ws):
    """Create Evidence Register sheet"""
    
    # Title
    ws.merge_cells('A1:N1')
    ws['A1'] = "Discovery Evidence Register"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Evidence ID', 15),
        ('B', 'Asset Category', 25),
        ('C', 'Discovery Method', 30),
        ('D', 'Evidence Type', 30),
        ('E', 'Evidence Description', 50),
        ('F', 'Evidence Location', 40),
        ('G', 'Collection Date', 15),
        ('H', 'Collected By', 25),
        ('I', 'Validity Period', 20),
        ('J', 'Review Date', 15),
        ('K', 'Reviewed By', 25),
        ('L', 'Review Status', 20),
        ('M', 'Retention End Date', 18),
        ('N', 'Notes', 40),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Add sample rows (user will add actual evidence)
    sample_data = [
        ("DISC-001", "Information Assets", "Database Query", "SQL Query Results", "List of production databases from CMDB query", "/evidence/database_inventory_20260122.xlsx", "22.01.2026"),
        ("DISC-002", "IT Infrastructure", "Network Scan (Automated)", "Nmap Scan Results", "Network scan of 10.0.0.0/8 subnet", "/evidence/nmap_scan_20260122.xml", "22.01.2026"),
        ("DISC-003", "Applications", "SaaS Audit", "SaaS License Report", "Okta application report - all SSO-integrated apps", "/evidence/okta_apps_20260122.csv", "22.01.2026"),
        ("DISC-004", "Personnel Assets", "HR System Export", "HR Data Export", "Active employee list with roles and departments", "/evidence/hr_export_20260122.xlsx", "22.01.2026"),
    ]
    
    row = 4
    for evidence_id, category, method, evidence_type, description, location, date in sample_data:
        ws[f'A{row}'] = evidence_id
        ws[f'B{row}'] = category
        ws[f'C{row}'] = method
        ws[f'D{row}'] = evidence_type
        ws[f'E{row}'] = description
        ws[f'F{row}'] = location
        ws[f'G{row}'] = date
        
        # Unlock cells for user input
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Add empty rows for additional evidence
    for i in range(20):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    # Data validations
    asset_categories = ["Information Assets", "IT Infrastructure", "Applications", "Physical Assets", "Personnel Assets", "All Categories"]
    dv_category = DataValidation(type="list", formula1=f'"{",".join(asset_categories)}"', allow_blank=True)
    dv_category.add(f'B4:B100')
    ws.add_data_validation(dv_category)
    
    dv_method = DataValidation(type="list", formula1=f'"{",".join(CONFIG["discovery_methods"])}"', allow_blank=True)
    dv_method.add(f'C4:C100')
    ws.add_data_validation(dv_method)
    
    review_statuses = ["Pending Review", "Reviewed - Valid", "Reviewed - Update Needed", "Reviewed - Invalid"]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(review_statuses)}"', allow_blank=True)
    dv_status.add(f'L4:L100')
    ws.add_data_validation(dv_status)


# Execute main function


def create_summary_dashboard_sheet(ws):
    """Create Summary Dashboard sheet"""
    CHECK = "✅"
    WARNING = "⚠️"
    XMARK = "❌"
    TARGET = "🎯"
    CHART = "📊"
    
    metrics_sheet = "Discovery Metrics & Summary"
    assessment_name = "Asset Discovery Assessment"
    compliance_ref = "D4"
    key_metrics = [
        ('Information Assets Discovered', 'B4', '100%', 'number'),
        ('IT Infrastructure Discovered', 'B5', '100%', 'number'),
        ('Applications Discovered', 'B6', '100%', 'number'),
        ('Physical Assets Discovered', 'B7', '100%', 'number'),
        ('Personnel Assets Discovered', 'B8', '100%', 'number'),
    ]
    
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{CHART} {assessment_name.upper()} - SUMMARY DASHBOARD"
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1F4E78', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:F2')
    ws['A2'] = "Quick overview of key metrics and compliance status"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].fill = PatternFill(start_color='E7E6E6', fill_type='solid')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    row = 4
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"{TARGET} OVERALL COMPLIANCE"
    ws[f'A{row}'].font = Font(size=13, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='2E75B5', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Overall Compliance:"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = f"='{metrics_sheet}'!{compliance_ref}"
    ws[f'D{row}'].font = Font(size=18, bold=True)
    ws[f'D{row}'].number_format = '0.0"%"'
    ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"{CHART} KEY METRICS"
    ws[f'A{row}'].font = Font(size=13, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='2E75B5', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    for col_idx, header in enumerate(["Metric", "Current", "Target", "Status"], start=1):
        ws.cell(row=row, column=col_idx, value=header)
        ws.cell(row=row, column=col_idx).font = Font(size=11, bold=True, color='FFFFFF')
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color='4472C4', fill_type='solid')
    row += 1
    
    for metric_label, cell_ref, target, format_type in key_metrics:
        ws.cell(row=row, column=1, value=metric_label)
        ws.cell(row=row, column=2, value=f"='{metrics_sheet}'!{cell_ref}")
        if format_type == 'number':
            ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3, value=target)
        ws.cell(row=row, column=4, value=f'=IF(B{row}>0,"{CHECK}","{XMARK}")')
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_approval_signoff_sheet(ws):
    """Create Approval & Sign-Off sheet"""
    CHECK = "✅"
    CLOCK = "⏳"
    XMARK = "❌"
    
    assessment_name = "Asset Discovery Assessment"
    checklist_items = [
        'All asset categories scanned/assessed',
        'Information assets documented completely',
        'IT infrastructure assets documented completely',
        'Applications inventory completed',
        'Physical assets inventory completed',
        'Personnel assets inventory completed',
        'Discovery metrics calculated and reviewed',
        'Evidence collected and registered',
        'Quality review completed',
        'Gap analysis performed',
        'Assessment reviewed by Information Security'
    ]
    
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{assessment_name.upper()} - APPROVAL & SIGN-OFF"
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1F4E78', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:F2')
    ws['A2'] = "Complete when assessment is finished and ready for approval"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    row = 4
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "COMPLETION CHECKLIST"
    ws[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='2E75B5', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    for col_idx, header in enumerate(["Item", "Status", "Completed By", "Date", "Notes"], start=1):
        ws.cell(row=row, column=col_idx, value=header)
        ws.cell(row=row, column=col_idx).font = Font(size=11, bold=True, color='FFFFFF')
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color='4472C4', fill_type='solid')
    row += 1
    
    checklist_start = row
    for item in checklist_items:
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=1).font = Font(size=10)
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFF2CC', fill_type='solid')
            ws.cell(row=row, column=col).protection = Protection(locked=False)
        row += 1
    
    from openpyxl.worksheet.datavalidation import DataValidation
    status_options = [f"{CHECK} Complete", f"{CLOCK} In Progress", f"{XMARK} Not Done"]
    dv = DataValidation(type="list", formula1=f'"{",".join(status_options)}"', allow_blank=True)
    dv.add(f'B{checklist_start}:B{row-1}')
    ws.add_data_validation(dv)
    
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "APPROVALS"
    ws[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='2E75B5', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    for col_idx, header in enumerate(["Role", "Name", "Signature", "Date", "Comments"], start=1):
        ws.cell(row=row, column=col_idx, value=header)
        ws.cell(row=row, column=col_idx).font = Font(size=11, bold=True, color='FFFFFF')
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color='4472C4', fill_type='solid')
    row += 1
    
    for approver in ["Asset Management Lead", "Information Security Manager", "CISO"]:
        ws.cell(row=row, column=1, value=approver)
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E1F2', fill_type='solid')
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFF2CC', fill_type='solid')
            ws.cell(row=row, column=col).protection = Protection(locked=False)
        row += 1
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30



if __name__ == "__main__":
    main()
