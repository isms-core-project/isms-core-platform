#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ISMS-IMP-A.5.23.S5 - Compliance Monitoring Dashboard Generator (REGULATORY)
ISO/IEC 27001:2022 Control A.5.23 + DORA/NIS2/AI Act/CLOUD Act

CRITICAL: This script consolidates data from FOUR assessment workbooks.
Before processing, it validates the structure of each source workbook.

Requirements:
    pip install openpyxl

Usage:
    python3 generate_reg_a523_5_compliance_dashboard.py
    
Inputs (must exist in current directory):
    - ISMS-IMP-A.5.23.S1_Inventory_*.xlsx  (from generate_reg_a523_1_inventory.py)
    - ISMS-IMP-A.5.23.S2_VendorDD_*.xlsx   (from generate_reg_a523_2_vendor_dd.py)
    - ISMS-IMP-A.5.23.S3_SecureConfig_*.xlsx (from generate_reg_a523_3_secure_config.py)
    - ISMS-IMP-A.5.23.S4_Governance_*.xlsx (from generate_reg_a523_4_governance.py)
    
Output:
    ISMS-IMP-A.5.23.S5_Dashboard_YYYYMMDD.xlsx

"What I cannot create, I do not understand." - Richard Feynman
"""

from datetime import datetime
import os
import glob
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# UNICODE SYMBOLS - UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
CLOUD = '\u2601'      # ☁️  Cloud
GLOBE = '\U0001F310'  # 🌐 Globe
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow


# ============================================================================
# SECTION 1: WORKBOOK SCHEMA DEFINITIONS
# ============================================================================

# Define the EXPECTED structure of each input workbook
# This prevents assumptions and enables schema validation

INVENTORY_SCHEMA = {
    'filename_pattern': 'ISMS-IMP-A.5.23.S1_Inventory_*.xlsx',
    'required_sheets': [
        'Instructions & Legend',
        '2. SaaS Services',
        '3. IaaS PaaS Services',
        '4. Cloud Security Services',  # This is Exit Strategy sheet
        '5. Cloud Storage Services',
    ],
    'exit_strategy_sheet': '4. Cloud Security Services',
    'exit_strategy_columns': {
        'Service Name': 'A',
        'Provider': 'B',
        'Service Type': 'C',
        'Criticality': 'H',
        'Exit Strategy Type': 'R',
        'Alternative Identified': 'S',
        'Export Tested': 'U',
        'Migration Complexity': 'W',
        'Lock-In Risk': 'X',
    },
    'data_start_row': 4,
}

VENDOR_DD_SCHEMA = {
    'filename_pattern': 'ISMS-IMP-A.5.23.S2_VendorDD_*.xlsx',
    'required_sheets': ['Instructions & Legend'],
}

SECURE_CONFIG_SCHEMA = {
    'filename_pattern': 'ISMS-IMP-A.5.23.S3_SecureConfig_*.xlsx',
    'required_sheets': ['Instructions & Legend'],
}

GOVERNANCE_SCHEMA = {
    'filename_pattern': 'ISMS-IMP-A.5.23.S4_Governance_*.xlsx',
    'required_sheets': [
        'Instructions & Legend',
        '7. Exit Strategy Review',  # NEW in v2.1
    ],
    'exit_review_sheet': '7. Exit Strategy Review',
    'exit_review_columns': {
        'Service Name': 'A',
        'Criticality': 'C',
        'PoC Testing Required?': 'H',
        'PoC Test Date (Last)': 'I',
        'PoC Test Result': 'J',
        'PoC Test Next Due': 'K',
        'Review Status': 'G',
    },
    'data_start_row': 4,
}


# ============================================================================
# SECTION 2: WORKBOOK VALIDATION
# ============================================================================

def find_latest_workbook(pattern: str) -> str:
    """Find the most recent workbook matching pattern."""
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(f"No workbook found matching: {pattern}")
    
    # Sort by modification time, newest first
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]


def validate_workbook_schema(filepath: str, schema: dict) -> tuple:
    """
    Validate that workbook matches expected schema.
    
    Returns:
        (is_valid, errors) - tuple of bool and list of error messages
    """
    errors = []
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        errors.append(f"Cannot open workbook: {e}")
        return (False, errors)
    
    # Check required sheets exist
    for required_sheet in schema['required_sheets']:
        if required_sheet not in wb.sheetnames:
            errors.append(f"Missing required sheet: '{required_sheet}'")
    
    # If exit strategy sheet defined, check columns
    if 'exit_strategy_sheet' in schema:
        sheet_name = schema['exit_strategy_sheet']
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Check column headers in row 3
            for col_name, col_letter in schema['exit_strategy_columns'].items():
                cell_value = ws[f'{col_letter}3'].value
                if cell_value != col_name:
                    errors.append(
                        f"Sheet '{sheet_name}' column {col_letter}: "
                        f"Expected '{col_name}', found '{cell_value}'"
                    )
    
    # If exit review sheet defined, check columns
    if 'exit_review_sheet' in schema:
        sheet_name = schema['exit_review_sheet']
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col_name, col_letter in schema['exit_review_columns'].items():
                cell_value = ws[f'{col_letter}3'].value
                if cell_value != col_name:
                    errors.append(
                        f"Sheet '{sheet_name}' column {col_letter}: "
                        f"Expected '{col_name}', found '{cell_value}'"
                    )
    
    wb.close()
    return (len(errors) == 0, errors)


def validate_all_inputs() -> dict:
    """
    Validate all input workbooks before processing.
    
    Returns:
        dict mapping workbook type to filepath
    """
    print("\n" + "=" * 78)
    print("WORKBOOK VALIDATION")
    print("=" * 78)
    
    workbooks = {}
    
    schemas = {
        'inventory': INVENTORY_SCHEMA,
        'vendor_dd': VENDOR_DD_SCHEMA,
        'secure_config': SECURE_CONFIG_SCHEMA,
        'governance': GOVERNANCE_SCHEMA,
    }
    
    all_valid = True
    
    for wb_type, schema in schemas.items():
        print(f"\n[{wb_type.upper()}] Searching for: {schema['filename_pattern']}")
        
        try:
            filepath = find_latest_workbook(schema['filename_pattern'])
            print(f"   [FOUND] {filepath}")
            
            is_valid, errors = validate_workbook_schema(filepath, schema)
            
            if is_valid:
                print(f"   [OK] Schema validation passed")
                workbooks[wb_type] = filepath
            else:
                print(f"   [ERROR] Schema validation failed:")
                for error in errors:
                    print(f"      - {error}")
                all_valid = False
        
        except FileNotFoundError as e:
            print(f"   [ERROR] {e}")
            all_valid = False
    
    if not all_valid:
        raise ValueError("Workbook validation failed. Fix errors and re-run.")
    
    print("\n" + "=" * 78)
    print("[OK] All workbooks validated successfully")
    print("=" * 78)
    
    return workbooks


# ============================================================================
# SECTION 3: DATA EXTRACTION (CONTROL-SPECIFIC)
# ============================================================================

def extract_exit_strategy_data(inventory_filepath: str) -> dict:
    """
    Extract Exit Strategy assessment data from Inventory workbook.
    
    Returns:
        dict with exit strategy metrics
    """
    print("\n[EXTRACT] Reading Exit Strategy data from inventory workbook...")
    
    wb = load_workbook(inventory_filepath, read_only=True, data_only=True)
    ws = wb[INVENTORY_SCHEMA['exit_strategy_sheet']]
    
    data = {
        'total_services': 0,
        'cloud_to_cloud': 0,
        'hybrid': 0,
        'on_premises': 0,
        'not_determined': 0,
        'high_lock_in': 0,
        'critical_lock_in': 0,
        'export_not_tested_critical': 0,
        'services': []
    }
    
    cols = INVENTORY_SCHEMA['exit_strategy_columns']
    start_row = INVENTORY_SCHEMA['data_start_row']
    
    for row in range(start_row, ws.max_row + 1):
        service_name = ws[f"{cols['Service Name']}{row}"].value
        
        if not service_name or str(service_name).strip() == "":
            continue  # Skip empty rows
        
        data['total_services'] += 1
        
        exit_strategy = ws[f"{cols['Exit Strategy Type']}{row}"].value
        criticality = ws[f"{cols['Criticality']}{row}"].value
        lock_in_risk = ws[f"{cols['Lock-In Risk']}{row}"].value
        export_tested = ws[f"{cols['Export Tested']}{row}"].value
        
        # Count by exit strategy type
        if exit_strategy == "Cloud-to-Cloud":
            data['cloud_to_cloud'] += 1
        elif exit_strategy == "Hybrid":
            data['hybrid'] += 1
        elif exit_strategy == "On-Premises":
            data['on_premises'] += 1
        else:
            data['not_determined'] += 1
        
        # Count lock-in risks
        if lock_in_risk == "High":
            data['high_lock_in'] += 1
        elif lock_in_risk == "Critical":
            data['critical_lock_in'] += 1
        
        # Count critical services without tested export
        if criticality == "Critical" and export_tested == "No":
            data['export_not_tested_critical'] += 1
        
        # Store individual service data
        data['services'].append({
            'name': service_name,
            'provider': ws[f"{cols['Provider']}{row}"].value,
            'type': ws[f"{cols['Service Type']}{row}"].value,
            'criticality': criticality,
            'exit_strategy': exit_strategy,
            'lock_in_risk': lock_in_risk,
            'export_tested': export_tested,
        })
    
    wb.close()
    
    print(f"   [OK] Extracted data from {data['total_services']} services")
    print(f"        Cloud-to-Cloud: {data['cloud_to_cloud']}")
    print(f"        Hybrid: {data['hybrid']}")
    print(f"        On-Premises: {data['on_premises']}")
    print(f"        Not Determined: {data['not_determined']}")
    
    return data


def extract_poc_testing_data(governance_filepath: str) -> dict:
    """
    Extract PoC testing compliance data from Governance workbook.
    
    Returns:
        dict with PoC testing metrics
    """
    print("\n[EXTRACT] Reading PoC Testing data from governance workbook...")
    
    wb = load_workbook(governance_filepath, read_only=True, data_only=True)
    
    # Check if Exit Strategy Review sheet exists
    if GOVERNANCE_SCHEMA['exit_review_sheet'] not in wb.sheetnames:
        print("   [WARNING] No Exit Strategy Review sheet found in governance workbook")
        wb.close()
        return {
            'total_requiring_poc': 0,
            'poc_completed': 0,
            'poc_overdue': 0,
            'poc_not_tested': 0,
            'services': []
        }
    
    ws = wb[GOVERNANCE_SCHEMA['exit_review_sheet']]
    
    data = {
        'total_requiring_poc': 0,
        'poc_completed': 0,
        'poc_overdue': 0,
        'poc_not_tested': 0,
        'services': []
    }
    
    cols = GOVERNANCE_SCHEMA['exit_review_columns']
    start_row = GOVERNANCE_SCHEMA['data_start_row']
    
    today = datetime.now().date()
    
    for row in range(start_row, ws.max_row + 1):
        service_name = ws[f"{cols['Service Name']}{row}"].value
        
        if not service_name or str(service_name).strip() == "":
            continue
        
        poc_required = ws[f"{cols['PoC Testing Required?']}{row}"].value
        poc_result = ws[f"{cols['PoC Test Result']}{row}"].value
        poc_next_due = ws[f"{cols['PoC Test Next Due']}{row}"].value
        
        if poc_required == "Yes (Critical)":
            data['total_requiring_poc'] += 1
            
            if poc_result == "Pass":
                data['poc_completed'] += 1
            elif poc_result in ["Not Tested"]:
                data['poc_not_tested'] += 1
            elif poc_result in ["In Progress"]:
                data['poc_overdue'] += 1  # Treat In Progress as potentially overdue
            
            # Check if actually overdue based on date
            if poc_next_due:
                try:
                    if isinstance(poc_next_due, datetime):
                        due_date = poc_next_due.date()
                    else:
                        due_date = datetime.strptime(str(poc_next_due), '%Y-%m-%d').date()
                    
                    if due_date < today and poc_result != "Pass":
                        data['poc_overdue'] += 1
                except:
                    pass  # Ignore date parsing errors
            
            data['services'].append({
                'name': service_name,
                'poc_result': poc_result,
                'criticality': ws[f"{cols['Criticality']}{row}"].value,
            })
    
    wb.close()
    
    print(f"   [OK] Extracted PoC testing data from {data['total_requiring_poc']} Critical services")
    print(f"        Completed (Pass): {data['poc_completed']}")
    print(f"        Not Tested: {data['poc_not_tested']}")
    print(f"        Overdue: {data['poc_overdue']}")
    
    return data


# ============================================================================
# SECTION 4: STYLE DEFINITIONS
# ============================================================================

def setup_styles() -> dict:
    """Define all cell styles (reuse from other generators)."""
    return {
        "header": {
            "font_params": {"name": "Calibri", "size": 14, "bold": True, "color": "FFFFFF"},
            "fill_params": {"start_color": "003366", "end_color": "003366", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "subheader": {
            "font_params": {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"},
            "fill_params": {"start_color": "4472C4", "end_color": "4472C4", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "column_header": {
            "font_params": {"name": "Calibri", "size": 10, "bold": True},
            "fill_params": {"start_color": "D9D9D9", "end_color": "D9D9D9", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "input_cell": {
            "fill_params": {"start_color": "FFFFCC", "end_color": "FFFFCC", "fill_type": "solid"},
            "alignment_params": {"horizontal": "left", "vertical": "center", "wrap_text": True},
        },
        "metric_good": {
            "fill_params": {"start_color": "C6EFCE", "end_color": "C6EFCE", "fill_type": "solid"},
            "font_params": {"name": "Calibri", "size": 11, "bold": True},
        },
        "metric_warning": {
            "fill_params": {"start_color": "FFEB9C", "end_color": "FFEB9C", "fill_type": "solid"},
            "font_params": {"name": "Calibri", "size": 11, "bold": True},
        },
        "metric_critical": {
            "fill_params": {"start_color": "FFC7CE", "end_color": "FFC7CE", "fill_type": "solid"},
            "font_params": {"name": "Calibri", "size": 11, "bold": True, "color": "9C0006"},
        },
    }


# ============================================================================
# SECTION 5: DASHBOARD CREATION
# ============================================================================

def create_exit_strategy_dashboard(ws, styles, exit_data):
    """Create Exit Strategy Coverage dashboard section."""
    
    # === HEADER ===
    ws.merge_cells('A1:H1')
    ws['A1'] = "EXIT STRATEGY COVERAGE DASHBOARD"
    ws['A1'].font = Font(**styles['header']['font_params'])
    ws['A1'].fill = PatternFill(**styles['header']['fill_params'])
    ws['A1'].alignment = Alignment(**styles['header']['alignment_params'])
    ws.row_dimensions[1].height = 40
    
    # === POLICY REFERENCE ===
    ws.merge_cells('A2:H2')
    ws['A2'] = "POL-S5 Section 8: Exit Strategy Framework | Cloud-to-Cloud (90%+), Hybrid (5-10%), On-Premises (<5%)"
    ws['A2'].font = Font(name="Calibri", size=9, italic=True)
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 25
    
    # === METRICS TABLE HEADER ===
    ws['A4'] = "METRIC"
    ws['B4'] = "VALUE"
    ws['C4'] = "PERCENTAGE"
    ws['D4'] = "STATUS"
    ws['E4'] = "POLICY TARGET"
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}4'].font = Font(bold=True)
        ws[f'{col}4'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f'{col}4'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    
    # === CALCULATE PERCENTAGES ===
    total = exit_data['total_services']
    if total > 0:
        c2c_pct = (exit_data['cloud_to_cloud'] / total) * 100
        hybrid_pct = (exit_data['hybrid'] / total) * 100
        onprem_pct = (exit_data['on_premises'] / total) * 100
        not_det_pct = (exit_data['not_determined'] / total) * 100
    else:
        c2c_pct = hybrid_pct = onprem_pct = not_det_pct = 0
    
    # === POPULATE METRICS ===
    metrics = [
        ("Total Cloud Services", total, "", "", "All services inventoried"),
        ("Cloud-to-Cloud Strategy", exit_data['cloud_to_cloud'], f"{c2c_pct:.1f}%", 
         CHECK if c2c_pct >= 85 else WARNING, "≥90% (default strategy)"),
        ("Hybrid Strategy", exit_data['hybrid'], f"{hybrid_pct:.1f}%", 
         CHECK if hybrid_pct <= 15 else WARNING, "5-10% (data sovereignty cases)"),
        ("On-Premises Strategy", exit_data['on_premises'], f"{onprem_pct:.1f}%", 
         CHECK if onprem_pct <= 5 else XMARK, "<5% (requires TCO justification)"),
        ("Not Yet Determined", exit_data['not_determined'], f"{not_det_pct:.1f}%", 
         CHECK if not_det_pct <= 10 else WARNING, "Temporary (new services only)"),
        ("", "", "", "", ""),
        ("High Lock-In Risk", exit_data['high_lock_in'], "", 
         WARNING if exit_data['high_lock_in'] > 0 else CHECK, "Mitigation plan required"),
        ("CRITICAL Lock-In Risk", exit_data['critical_lock_in'], "", 
         XMARK if exit_data['critical_lock_in'] > 0 else CHECK, "Immediate action required"),
        ("Export Not Tested (Critical)", exit_data['export_not_tested_critical'], "", 
         XMARK if exit_data['export_not_tested_critical'] > 0 else CHECK, "All Critical services tested"),
    ]
    
    row = 5
    for metric, value, pct, status, target in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        ws[f'C{row}'] = pct
        ws[f'D{row}'] = status
        ws[f'E{row}'] = target
        
        # Conditional coloring for On-Premises warning
        if metric == "On-Premises Strategy" and onprem_pct > 5:
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].fill = PatternFill(**styles['metric_critical']['fill_params'])
                ws[f'{col}{row}'].font = Font(**styles['metric_critical']['font_params'])
        
        # Conditional coloring for Critical Lock-In
        if metric == "CRITICAL Lock-In Risk" and exit_data['critical_lock_in'] > 0:
            for col in ['A', 'B']:
                ws[f'{col}{row}'].fill = PatternFill(**styles['metric_critical']['fill_params'])
                ws[f'{col}{row}'].font = Font(**styles['metric_critical']['font_params'])
        
        row += 1
    
    print("   [OK] Exit Strategy dashboard section created")


def create_poc_testing_dashboard(ws, styles, poc_data, start_row):
    """Create PoC Testing Compliance dashboard section."""
    
    # === HEADER ===
    ws.merge_cells(f'A{start_row}:H{start_row}')
    ws[f'A{start_row}'] = "DORA PoC TESTING COMPLIANCE (Article 28.6)"
    ws[f'A{start_row}'].font = Font(**styles['subheader']['font_params'])
    ws[f'A{start_row}'].fill = PatternFill(**styles['subheader']['fill_params'])
    ws[f'A{start_row}'].alignment = Alignment(**styles['subheader']['alignment_params'])
    ws.row_dimensions[start_row].height = 35
    
    start_row += 1
    
    # === POLICY REFERENCE ===
    ws.merge_cells(f'A{start_row}:H{start_row}')
    ws[f'A{start_row}'] = "DORA Article 28.6: Financial entities must test exit strategies annually for critical ICT providers"
    ws[f'A{start_row}'].font = Font(name="Calibri", size=9, italic=True)
    ws[f'A{start_row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[start_row].height = 25
    
    start_row += 2
    
    # === METRICS TABLE HEADER ===
    ws[f'A{start_row}'] = "METRIC"
    ws[f'B{start_row}'] = "VALUE"
    ws[f'C{start_row}'] = "STATUS"
    ws[f'D{start_row}'] = "COMPLIANCE REQUIREMENT"
    
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{start_row}'].font = Font(bold=True)
        ws[f'{col}{start_row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f'{col}{start_row}'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.column_dimensions['D'].width = 35
    
    start_row += 1
    
    # === POPULATE METRICS ===
    total_requiring = poc_data['total_requiring_poc']
    completed = poc_data['poc_completed']
    not_tested = poc_data['poc_not_tested']
    overdue = poc_data['poc_overdue']
    
    # Calculate compliance percentage
    if total_requiring > 0:
        compliance_pct = (completed / total_requiring) * 100
    else:
        compliance_pct = 100  # No services require testing = compliant
    
    metrics = [
        ("Critical Services Requiring PoC Testing", total_requiring, "", "All Critical services with exit strategy"),
        ("PoC Testing Completed (Pass)", completed, 
         CHECK if completed >= total_requiring else WARNING, "100% annually"),
        ("PoC Testing Not Tested", not_tested, 
         XMARK if not_tested > 0 else CHECK, "0 (all must be tested)"),
        ("PoC Testing Overdue", overdue, 
         XMARK if overdue > 0 else CHECK, "0 (12-month cycle)"),
        ("Overall Compliance", f"{compliance_pct:.1f}%", 
         CHECK if compliance_pct == 100 else XMARK, "100% (DORA requirement)"),
    ]
    
    for metric, value, status, requirement in metrics:
        ws[f'A{start_row}'] = metric
        ws[f'B{start_row}'] = value
        ws[f'C{start_row}'] = status
        ws[f'D{start_row}'] = requirement
        
        # Highlight overdue PoC testing
        if metric == "PoC Testing Overdue" and overdue > 0:
            for col in ['A', 'B', 'C']:
                ws[f'{col}{start_row}'].fill = PatternFill(**styles['metric_critical']['fill_params'])
                ws[f'{col}{start_row}'].font = Font(**styles['metric_critical']['font_params'])
        
        # Highlight not tested
        if metric == "PoC Testing Not Tested" and not_tested > 0:
            for col in ['A', 'B', 'C']:
                ws[f'{col}{start_row}'].fill = PatternFill(**styles['metric_warning']['fill_params'])
                ws[f'{col}{start_row}'].font = Font(**styles['metric_warning']['font_params'])
        
        start_row += 1
    
    print("   [OK] PoC Testing dashboard section created")


def create_dashboard_workbook(workbooks: dict) -> Workbook:
    """Create consolidated dashboard workbook."""
    
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    wb.create_sheet(title="Exit Strategy Dashboard")
    wb.create_sheet(title="Risk Overview")
    wb.create_sheet(title="Recommendations")
    
    styles = setup_styles()
    
    # === EXTRACT DATA FROM SOURCE WORKBOOKS ===
    print("\n" + "=" * 78)
    print("DATA EXTRACTION")
    print("=" * 78)
    
    exit_data = extract_exit_strategy_data(workbooks['inventory'])
    poc_data = extract_poc_testing_data(workbooks['governance'])
    
    # === CREATE DASHBOARD SHEETS ===
    print("\n" + "=" * 78)
    print("DASHBOARD CREATION")
    print("=" * 78)
    
    print("\n[CREATE] Building Exit Strategy Dashboard...")
    ws1 = wb["Exit Strategy Dashboard"]
    create_exit_strategy_dashboard(ws1, styles, exit_data)
    create_poc_testing_dashboard(ws1, styles, poc_data, start_row=20)
    
    # === RISK OVERVIEW SHEET ===
    print("\n[CREATE] Building Risk Overview...")
    ws2 = wb["Risk Overview"]
    
    ws2.merge_cells('A1:F1')
    ws2['A1'] = "EXIT STRATEGY RISK OVERVIEW"
    ws2['A1'].font = Font(**styles['header']['font_params'])
    ws2['A1'].fill = PatternFill(**styles['header']['fill_params'])
    ws2['A1'].alignment = Alignment(**styles['header']['alignment_params'])
    ws2.row_dimensions[1].height = 40
    
    # High Risk Services Table
    ws2['A3'] = "HIGH RISK SERVICES"
    ws2['A3'].font = Font(bold=True, size=12)
    
    ws2['A4'] = "Service Name"
    ws2['B4'] = "Provider"
    ws2['C4'] = "Criticality"
    ws2['D4'] = "Exit Strategy"
    ws2['E4'] = "Lock-In Risk"
    ws2['F4'] = "Export Tested"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws2[f'{col}4'].font = Font(bold=True)
        ws2[f'{col}4'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 15
    
    # Populate high-risk services
    row = 5
    for service in exit_data['services']:
        # Flag high-risk services
        is_high_risk = (
            service['lock_in_risk'] in ['High', 'Critical'] or
            (service['criticality'] == 'Critical' and service['export_tested'] == 'No') or
            service['exit_strategy'] == 'On-Premises'
        )
        
        if is_high_risk:
            ws2[f'A{row}'] = service['name']
            ws2[f'B{row}'] = service['provider']
            ws2[f'C{row}'] = service['criticality']
            ws2[f'D{row}'] = service['exit_strategy']
            ws2[f'E{row}'] = service['lock_in_risk']
            ws2[f'F{row}'] = service['export_tested']
            
            # Highlight row based on risk
            if service['lock_in_risk'] == 'Critical':
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws2[f'{col}{row}'].fill = PatternFill(**styles['metric_critical']['fill_params'])
            
            row += 1
    
    print("   [OK] Risk Overview sheet created")
    
    # === RECOMMENDATIONS SHEET ===
    print("\n[CREATE] Building Recommendations...")
    ws3 = wb["Recommendations"]
    
    ws3.merge_cells('A1:D1')
    ws3['A1'] = "EXIT STRATEGY RECOMMENDATIONS"
    ws3['A1'].font = Font(**styles['header']['font_params'])
    ws3['A1'].fill = PatternFill(**styles['header']['fill_params'])
    ws3['A1'].alignment = Alignment(**styles['header']['alignment_params'])
    ws3.row_dimensions[1].height = 40
    
    ws3.column_dimensions['A'].width = 50
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 20
    
    # Generate recommendations based on findings
    recommendations = []
    
    # On-Premises warnings
    if exit_data['on_premises'] > 0:
        recommendations.append((
            XMARK,
            "CRITICAL",
            f"On-Premises exit strategy selected for {exit_data['on_premises']} service(s)",
            "Validate TCO analysis with CISO. On-Premises justified in <5% of cases (POL-S5 Section 8.1.1.3)."
        ))
    
    # Critical lock-in
    if exit_data['critical_lock_in'] > 0:
        recommendations.append((
            XMARK,
            "HIGH",
            f"Critical lock-in risk for {exit_data['critical_lock_in']} service(s)",
            "Develop vendor-independent architecture roadmap. Consider multi-cloud or abstraction layers."
        ))
    
    # Export not tested
    if exit_data['export_not_tested_critical'] > 0:
        recommendations.append((
            WARNING,
            "MEDIUM",
            f"Export not tested for {exit_data['export_not_tested_critical']} Critical service(s)",
            "Test data export capability immediately. DORA Art. 28.6 requires annual PoC testing."
        ))
    
    # PoC testing overdue
    if poc_data['poc_overdue'] > 0:
        recommendations.append((
            XMARK,
            "CRITICAL",
            f"PoC testing overdue for {poc_data['poc_overdue']} Critical service(s)",
            "DORA Art. 28.6 VIOLATION. Conduct PoC testing immediately and report to CISO."
        ))
    
    # Good compliance
    if not recommendations:
        recommendations.append((
            CHECK,
            "GOOD",
            "Exit Strategy framework fully compliant",
            "Continue annual reviews and PoC testing. Monitor for new services."
        ))
    
    # Populate recommendations
    ws3['A3'] = "Priority"
    ws3['B3'] = "Severity"
    ws3['C3'] = "Finding"
    ws3['D3'] = "Recommendation"
    
    for col in ['A', 'B', 'C', 'D']:
        ws3[f'{col}3'].font = Font(bold=True)
        ws3[f'{col}3'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    ws3.column_dimensions['C'].width = 45
    ws3.column_dimensions['D'].width = 60
    
    row = 4
    for icon, severity, finding, recommendation in recommendations:
        ws3[f'A{row}'] = icon
        ws3[f'B{row}'] = severity
        ws3[f'C{row}'] = finding
        ws3[f'D{row}'] = recommendation
        
        # Color code by severity
        if severity == "CRITICAL":
            for col in ['A', 'B']:
                ws3[f'{col}{row}'].fill = PatternFill(**styles['metric_critical']['fill_params'])
                ws3[f'{col}{row}'].font = Font(**styles['metric_critical']['font_params'])
        elif severity == "HIGH":
            for col in ['A', 'B']:
                ws3[f'{col}{row}'].fill = PatternFill(**styles['metric_warning']['fill_params'])
                ws3[f'{col}{row}'].font = Font(**styles['metric_warning']['font_params'])
        elif severity == "GOOD":
            for col in ['A', 'B']:
                ws3[f'{col}{row}'].fill = PatternFill(**styles['metric_good']['fill_params'])
                ws3[f'{col}{row}'].font = Font(**styles['metric_good']['font_params'])
        
        ws3[f'C{row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws3[f'D{row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws3.row_dimensions[row].height = 30
        
        row += 1
    
    print("   [OK] Recommendations sheet created")
    
    return wb


# ============================================================================
# SECTION 6: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution - validates inputs and generates dashboard."""
    
    print("=" * 78)
    print("ISMS-IMP-A.5.23.S5 - Compliance Monitoring Dashboard Generator (REGULATORY)")
    print("ISO/IEC 27001:2022 Control A.5.23 + Exit Strategy Framework")
    print("=" * 78)
    print()
    print(f"{SHIELD} Consolidates data from 4 assessment workbooks:")
    print(f"   1. Inventory (Exit Strategy framework)")
    print(f"   2. Vendor Due Diligence")
    print(f"   3. Secure Configuration")
    print(f"   4. Governance (Exit Strategy review + PoC testing)")
    
    # STEP 1: Validate all input workbooks
    try:
        workbooks = validate_all_inputs()
    except ValueError as e:
        print(f"\n{XMARK} VALIDATION FAILED")
        print(f"   {str(e)}")
        return None
    
    # STEP 2: Create dashboard workbook
    print("\n" + "=" * 78)
    print("DASHBOARD GENERATION")
    print("=" * 78)
    
    wb = create_dashboard_workbook(workbooks)
    
    # STEP 3: Save workbook
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.5.23.S5_Dashboard_{timestamp}.xlsx"
    
    print(f"\n[SAVE] Saving dashboard: {filename}")
    wb.save(filename)
    
    print("\n" + "=" * 78)
    print(f"{CHECK} SUCCESS: {filename}")
    print("=" * 78)
    print()
    print(f"{CHART} DASHBOARD STRUCTURE (3 sheets):")
    print("   1. Exit Strategy Dashboard (Exit Strategy metrics + PoC Testing compliance)")
    print("   2. Risk Overview (High-risk services requiring attention)")
    print("   3. Recommendations (Actionable findings)")
    print()
    print(f"{SHIELD} COMPLIANCE COVERAGE:")
    print("   ✓ ISO 27001:2022 A.5.23 (Exit planning requirements)")
    print("   ✓ POL-A.5.19-23-S5 Section 8 (Exit Strategy Framework)")
    print("   ✓ DORA Article 28.6 (Annual PoC testing)")
    print()
    print(f"{TARGET} NEXT STEPS:")
    print("   1. Review Exit Strategy Dashboard for compliance status")
    print("   2. Address Critical/High priority items in Recommendations")
    print("   3. Escalate PoC testing failures to CISO immediately")
    print("   4. Update risk register with findings")
    print("   5. Quarterly reporting to Risk Committee")
    print()
    print("=" * 78)
    print('"In God we trust. All others must bring data." – W. Edwards Deming')
    print("=" * 78)
    
    return filename


if __name__ == "__main__":
    main()
