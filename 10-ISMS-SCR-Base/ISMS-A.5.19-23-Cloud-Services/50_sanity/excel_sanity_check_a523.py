#!/usr/bin/env python3
"""
excel_sanity_check_a523.py - Cloud Services Workbook Validator
ISO/IEC 27001:2022 Control A.5.23

Validates all 5 assessment workbooks:
- 5.23.1: Cloud Service Inventory & Classification
- 5.23.2: Vendor Due Diligence & Contracts
- 5.23.3: Secure Configuration & Deployment
- 5.23.4: Ongoing Governance & Risk Management
- 5.23.5: Compliance Monitoring & Exit Planning

Usage:
    python3 excel_sanity_check_a523.py <workbook.xlsx>
    python3 excel_sanity_check_a523.py --all

Returns:
    Exit code 0 if validation passes
    Exit code 1 if validation fails
"""

import sys
import os
import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ============================================================================
# WORKBOOK SPECIFICATIONS
# ============================================================================

WORKBOOK_SPECS = {
    # ------------------------------------------------------------------------
    # 5.23.1 - Cloud Service Inventory
    # ------------------------------------------------------------------------
    "5.23.1": {
        "name": "Cloud Service Inventory & Classification",
        "filename_pattern": "ISMS-IMP-A.5.23.S1_Inventory",
        "evidence_prefix": "EV-INV",
        "expected_sheets": [
            "Instructions & Legend",
            "2. SaaS Services",
            "3. IaaS PaaS Services",
            "4. Cloud Security Services",
            "5. Cloud Storage Services",
            "6. Data Classification Mapping",
            "7. Service Criticality",
            "8. Summary Dashboard",
            "9. Evidence Register",
            "10. Approval Sign-Off",
        ],
        "assessment_sheets": [
            "2. SaaS Services",
            "3. IaaS PaaS Services",
            "4. Cloud Security Services",
            "5. Cloud Storage Services",
        ],
        "header_row": 6,
        "data_entry_start_row": 8,
        "data_entry_end_row": 27,
        "base_column_headers": [
            "Cloud Service Name", "Service Type", "Vendor Name",
            "Service Criticality", "Data Classification", "Data Residency Region",
            "Contract Status", "Status",
        ],
        "dashboard_sheet": "8. Summary Dashboard",
        "evidence_sheet": "9. Evidence Register",
        "evidence_header_row": 3,
        "approval_sheet": "10. Approval Sign-Off",
        "approval_sections": ["ASSESSMENT COMPLETION", "REVIEWER APPROVAL",
                             "CISO APPROVAL", "NEXT REVIEW"],
    },
    # ------------------------------------------------------------------------
    # 5.23.2 - Vendor Due Diligence
    # ------------------------------------------------------------------------
    "5.23.2": {
        "name": "Vendor Due Diligence & Contracts",
        "filename_pattern": "ISMS-IMP-A.5.23.S2_VendorDD",
        "evidence_prefix": "EV-VDD",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Security Certifications",
            "3. Contract Terms",
            "4. SLA Performance",
            "5. Data Sovereignty",
            "6. Forensics & Audit",
            "7. Summary Dashboard",
            "8. Evidence Register",
            "9. Approval Sign-Off",
        ],
        "assessment_sheets": [
            "2. Security Certifications",
            "3. Contract Terms",
            "4. SLA Performance",
            "5. Data Sovereignty",
            "6. Forensics & Audit",
        ],
        "header_row": 4,
        "data_entry_start_row": 5,
        "data_entry_end_row": 24,
        "base_column_headers": [
            "Cloud Service Name", "Vendor Name", "Service Type",
            "Service Criticality", "Data Classification", "Contract Type",
            "Contract Start Date", "Status",
        ],
        "dashboard_sheet": "7. Summary Dashboard",
        "evidence_sheet": "8. Evidence Register",
        "evidence_header_row": 3,
        "approval_sheet": "9. Approval Sign-Off",
        "approval_sections": ["ASSESSMENT COMPLETION", "LEGAL REVIEW",
                             "PROCUREMENT APPROVAL", "CISO APPROVAL",
                             "NEXT REVIEW DETAILS"],
    },
    # ------------------------------------------------------------------------
    # 5.23.3 - Secure Configuration
    # ------------------------------------------------------------------------
    "5.23.3": {
        "name": "Secure Configuration & Deployment",
        "filename_pattern": "ISMS-IMP-A.5.23.S3_SecureConfig",
        "evidence_prefix": "EV-CFG",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Identity & Access",
            "3. Data Protection",
            "4. Network Security",
            "5. Logging & Monitoring",
            "6. Backup & Recovery",
            "7. Summary Dashboard",
            "8. Evidence Register",
            "9. Approval Sign-Off",
        ],
        "assessment_sheets": [
            "2. Identity & Access",
            "3. Data Protection",
            "4. Network Security",
            "5. Logging & Monitoring",
            "6. Backup & Recovery",
        ],
        "header_row": 4,
        "data_entry_start_row": 6,
        "data_entry_end_row": 24,
        "base_column_headers": [
            "Cloud Service Name", "Vendor Name", "Service Type",
            None, "Service Criticality", "Data Classification",
            None, "Status",
        ],
        "dashboard_sheet": "7. Summary Dashboard",
        "evidence_sheet": "8. Evidence Register",
        "evidence_header_row": 3,
        "approval_sheet": "9. Approval Sign-Off",
        "approval_sections": ["IT OPERATIONS", "SECURITY", "CISO"],
    },
    # ------------------------------------------------------------------------
    # 5.23.4 - Ongoing Governance
    # ------------------------------------------------------------------------
    "5.23.4": {
        "name": "Ongoing Governance & Risk Management",
        "filename_pattern": "ISMS-IMP-A.5.23.S4_Governance",
        "evidence_prefix": "EV-GOV",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Access Review",
            "3. Change Management",
            "4. Incident Management",
            "5. Business Continuity",
            "6. Vendor Risk Monitoring",
            "7. Summary Dashboard",
            "8. Evidence Register",
            "9. Approval Sign-Off",
        ],
        "assessment_sheets": [
            "2. Access Review",
            "3. Change Management",
            "4. Incident Management",
            "5. Business Continuity",
            "6. Vendor Risk Monitoring",
        ],
        "header_row": 4,
        "data_entry_start_row": 6,
        "data_entry_end_row": 24,
        "base_column_headers": [
            "Cloud Service Name", "Vendor Name", "Service Type",
            "Review Period", "Service Criticality", "Data Classification",
            None, "Status",
        ],
        "dashboard_sheet": "7. Summary Dashboard",
        "evidence_sheet": "8. Evidence Register",
        "evidence_header_row": 4,
        "approval_sheet": "9. Approval Sign-Off",
        "approval_sections": ["IT OPERATIONS", "COMPLIANCE", "CISO"],
    },
    # ------------------------------------------------------------------------
    # 5.23.5 - Compliance & Exit
    # ------------------------------------------------------------------------
    "5.23.5": {
        "name": "Compliance Monitoring & Exit Planning",
        "filename_pattern": "ISMS-IMP-A.5.23.S5_Compliance",
        "evidence_prefix": "EV-CMP",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Compliance Monitoring",
            "3. SLA Performance",
            "4. Exit Planning",
            "5. Data Portability",
            "6. Termination Readiness",
            "7. Summary Dashboard",
            "8. Evidence Register",
            "9. Approval Sign-Off",
        ],
        "assessment_sheets": [
            "2. Compliance Monitoring",
            "3. SLA Performance",
            "4. Exit Planning",
            "5. Data Portability",
            "6. Termination Readiness",
        ],
        "header_row": 4,
        "data_entry_start_row": 6,
        "data_entry_end_row": 24,
        "base_column_headers": [
            "Cloud Service Name", "Vendor Name", "Service Type",
            "Compliance Framework", "Service Criticality", "Data Classification",
            None, "Status",
        ],
        "dashboard_sheet": "7. Summary Dashboard",
        "evidence_sheet": "8. Evidence Register",
        "evidence_header_row": 4,
        "approval_sheet": "9. Approval Sign-Off",
        "approval_sections": ["COMPLIANCE", "LEGAL", "CISO"],
    },
}


# ============================================================================
# WORKBOOK DETECTION
# ============================================================================

def detect_workbook_type(filename):
    """Auto-detect workbook type from filename."""
    basename = os.path.basename(filename)
    
    for wb_type, spec in WORKBOOK_SPECS.items():
        if spec["filename_pattern"] in basename:
            return wb_type, spec
    
    # Fallback pattern matching
    fallbacks = [
        ("5.23.1", ["5.23.1", "Inventory"]),
        ("5.23.2", ["5.23.2", "VendorDD"]),
        ("5.23.3", ["5.23.3", "SecureConfig", "Config"]),
        ("5.23.4", ["5.23.4", "Governance"]),
        ("5.23.5", ["5.23.5", "Compliance", "Exit"]),
    ]
    
    for wb_type, patterns in fallbacks:
        if any(p in basename for p in patterns):
            return wb_type, WORKBOOK_SPECS[wb_type]
    
    return None, None


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_sheet_structure(wb, spec):
    """Validate workbook has expected sheets."""
    issues = []
    actual_sheets = wb.sheetnames
    expected_sheets = spec["expected_sheets"]
    
    if len(actual_sheets) != len(expected_sheets):
        issues.append(
            f"[X] Sheet count mismatch: Expected {len(expected_sheets)}, "
            f"found {len(actual_sheets)}"
        )
    
    for idx, expected_name in enumerate(expected_sheets):
        if idx >= len(actual_sheets):
            issues.append(f"[X] Missing sheet: '{expected_name}'")
            continue
        if actual_sheets[idx] != expected_name:
            issues.append(
                f"[X] Sheet {idx+1}: Expected '{expected_name}', "
                f"found '{actual_sheets[idx]}'"
            )
    
    return issues


def validate_assessment_columns(ws, spec, sheet_name):
    """Validate assessment sheet column headers."""
    issues = []
    header_row = spec.get("header_row", 4)
    expected_headers = spec.get("base_column_headers", [])
    
    for idx, expected in enumerate(expected_headers, start=1):
        if expected is None:  # Skip variable columns
            continue
        actual = ws.cell(row=header_row, column=idx).value
        if actual and expected.lower() not in str(actual).lower():
            issues.append(
                f"[X] [{sheet_name}] Col {get_column_letter(idx)}: "
                f"Expected '{expected}', found '{actual}'"
            )
    
    return issues


def validate_data_validations(ws, sheet_name):
    """Validate dropdown data validations exist."""
    issues = []
    
    if not hasattr(ws, 'data_validations'):
        issues.append(f"[X] [{sheet_name}] No data validations found")
        return issues
    
    validation_count = len(ws.data_validations.dataValidation)
    if validation_count < 3:
        issues.append(
            f"[!] [{sheet_name}] Low validation count: {validation_count} (expected >=3)"
        )
    
    return issues


def validate_dashboard(wb, spec):
    """Validate Summary Dashboard has formulas."""
    issues = []
    dashboard_name = spec.get("dashboard_sheet", "7. Summary Dashboard")
    
    if dashboard_name not in wb.sheetnames:
        issues.append(f"[X] Dashboard sheet '{dashboard_name}' not found")
        return issues
    
    ws = wb[dashboard_name]
    formula_found = False
    
    for row in range(3, 30):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                if cell.value.startswith("="):
                    formula_found = True
                    break
        if formula_found:
            break
    
    if not formula_found:
        issues.append("[X] [Dashboard] No formulas found - may not auto-calculate")
    
    return issues


def validate_evidence_register(wb, spec):
    """Validate Evidence Register structure."""
    issues = []
    evidence_name = spec.get("evidence_sheet", "8. Evidence Register")
    
    if evidence_name not in wb.sheetnames:
        issues.append(f"[X] Evidence Register '{evidence_name}' not found")
        return issues
    
    ws = wb[evidence_name]
    header_row = spec.get("evidence_header_row", 4)
    prefix = spec.get("evidence_prefix", "EV-")
    
    # Check for Evidence ID column header
    ev_id_header = ws.cell(row=header_row, column=1).value
    if ev_id_header and "Evidence" not in str(ev_id_header):
        issues.append(
            f"[X] [Evidence Register] Col A header: Expected 'Evidence ID', found '{ev_id_header}'"
        )
    
    # Check first data row for correct prefix
    first_data_row = header_row + 1
    first_id = ws.cell(row=first_data_row, column=1).value
    if first_id:
        if prefix not in str(first_id):
            issues.append(f"[!] [Evidence Register] ID prefix mismatch: expected '{prefix}', got '{first_id}'")
    
    return issues


def validate_approval_signoff(wb, spec):
    """Validate Approval Sign-Off sections."""
    issues = []
    approval_name = spec.get("approval_sheet", "9. Approval Sign-Off")
    
    if approval_name not in wb.sheetnames:
        issues.append(f"[X] Approval sheet '{approval_name}' not found")
        return issues
    
    ws = wb[approval_name]
    expected_sections = spec.get("approval_sections", [])
    found_sections = []
    
    # Scan first 60 rows for section headers
    content = ""
    for row in range(1, 60):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value:
            content += str(cell_value).upper() + " "
    
    for section in expected_sections:
        if section.upper() in content:
            found_sections.append(section)
    
    for section in expected_sections:
        if section not in found_sections:
            issues.append(f"[!] [Approval] Missing section: '{section}'")
    
    return issues


def validate_checklist_sections(wb, spec):
    """Validate checklist sections exist in assessment sheets."""
    issues = []
    
    for sheet_name in spec["assessment_sheets"]:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        checklist_found = False
        end_row = spec.get("data_entry_end_row", 24)
        
        # Scan rows after data area for "CHECKLIST"
        for row in range(end_row + 1, end_row + 30):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and "CHECKLIST" in str(cell_val).upper():
                checklist_found = True
                break
        
        if not checklist_found:
            issues.append(f"[!] [{sheet_name}] Checklist section not found")
    
    return issues


# ============================================================================
# MAIN VALIDATION
# ============================================================================

def validate_workbook(filename):
    """Main validation orchestrator."""
    
    print("=" * 70)
    print("EXCEL SANITY CHECK - Cloud Services Workbook Validator")
    print("ISO/IEC 27001:2022 Control A.5.23")
    print("=" * 70)
    print(f"\n[FILE] Workbook: {os.path.basename(filename)}")
    
    if not os.path.exists(filename):
        print(f"[X] ERROR: File not found: {filename}")
        return False
    
    wb_type, spec = detect_workbook_type(filename)
    if not wb_type:
        print("[X] ERROR: Cannot detect workbook type from filename")
        print("   Expected patterns:")
        for wt, sp in WORKBOOK_SPECS.items():
            print(f"     {wt}: {sp['filename_pattern']}*.xlsx")
        return False
    
    print(f"[TYPE] {wb_type} - {spec['name']}")
    print(f"[OK] Expected Sheets: {len(spec['expected_sheets'])}")
    
    try:
        wb = load_workbook(filename, data_only=False)
        print("[OK] Workbook loaded successfully")
    except Exception as e:
        print(f"[X] ERROR: Cannot load workbook: {e}")
        return False
    
    all_issues = []
    
    # 1. Sheet Structure
    print("\n[CHECK] Validating sheet structure...")
    issues = validate_sheet_structure(wb, spec)
    all_issues.extend(issues)
    print("   [OK]" if not issues else "\n".join(f"   {i}" for i in issues))
    
    # 2. Assessment Columns
    print("\n[CHECK] Validating assessment columns...")
    col_issues = []
    for sheet_name in spec["assessment_sheets"]:
        if sheet_name in wb.sheetnames:
            issues = validate_assessment_columns(wb[sheet_name], spec, sheet_name)
            col_issues.extend(issues)
    all_issues.extend(col_issues)
    if not col_issues:
        print(f"   [OK] All {len(spec['assessment_sheets'])} assessment sheets OK")
    else:
        for i in col_issues:
            print(f"   {i}")
    
    # 3. Data Validations
    print("\n[CHECK] Validating dropdowns...")
    dv_issues = []
    for sheet_name in spec["assessment_sheets"]:
        if sheet_name in wb.sheetnames:
            issues = validate_data_validations(wb[sheet_name], sheet_name)
            dv_issues.extend(issues)
    all_issues.extend(dv_issues)
    if not dv_issues:
        print(f"   [OK] Data validations present")
    else:
        for i in dv_issues:
            print(f"   {i}")
    
    # 4. Dashboard
    print("\n[CHECK] Validating dashboard...")
    issues = validate_dashboard(wb, spec)
    all_issues.extend(issues)
    print("   [OK]" if not issues else "\n".join(f"   {i}" for i in issues))
    
    # 5. Evidence Register
    print("\n[CHECK] Validating evidence register...")
    issues = validate_evidence_register(wb, spec)
    all_issues.extend(issues)
    print("   [OK]" if not issues else "\n".join(f"   {i}" for i in issues))
    
    # 6. Approval Sign-Off
    print("\n[CHECK] Validating approval sign-off...")
    issues = validate_approval_signoff(wb, spec)
    all_issues.extend(issues)
    print("   [OK]" if not issues else "\n".join(f"   {i}" for i in issues))
    
    # 7. Checklist Sections (for 5.23.3, 5.23.4, 5.23.5)
    if wb_type in ["5.23.3", "5.23.4", "5.23.5"]:
        print("\n[CHECK] Validating checklist sections...")
        issues = validate_checklist_sections(wb, spec)
        all_issues.extend(issues)
        print("   [OK]" if not issues else "\n".join(f"   {i}" for i in issues))
    
    wb.close()
    
    # Summary
    print("\n" + "=" * 70)
    error_count = len([i for i in all_issues if i.startswith("[X]")])
    warn_count = len([i for i in all_issues if i.startswith("[!]")])
    
    if not all_issues:
        print("[PASS] VALIDATION PASSED - NO ISSUES DETECTED")
        print("=" * 70)
        return True
    elif error_count == 0:
        print(f"[WARN] PASSED WITH {warn_count} WARNING(S)")
        print("=" * 70)
        return True
    else:
        print(f"[FAIL] VALIDATION FAILED - {error_count} error(s), {warn_count} warning(s)")
        print("=" * 70)
        return False


# ============================================================================
# BATCH MODE
# ============================================================================

def find_523_workbooks(directory="."):
    """Find all 5.23.x workbooks in directory."""
    patterns = [
        "ISMS-IMP-A.5.23.*.xlsx",
        "*_Inventory_*.xlsx",
        "*_VendorDD_*.xlsx",
        "*_SecureConfig_*.xlsx",
        "*_Governance_*.xlsx",
        "*_Compliance_*.xlsx",
    ]
    
    found = set()
    for pattern in patterns:
        matches = glob.glob(os.path.join(directory, pattern))
        found.update(matches)
    
    # Filter out temp files
    found = [f for f in found if not os.path.basename(f).startswith("~$")]
    
    return sorted(found)


def validate_all_workbooks(directory="."):
    """Validate all 5.23.x workbooks in directory."""
    workbooks = find_523_workbooks(directory)
    
    if not workbooks:
        print("\n[X] No 5.23.x workbooks found in current directory.")
        print("Looking for files matching: ISMS-IMP-A.5.23.*.xlsx")
        return False
    
    print(f"\n{'=' * 70}")
    print("BATCH VALIDATION - ISMS Control 5.23 Workbooks")
    print(f"Found {len(workbooks)} workbook(s)")
    print(f"{'=' * 70}")
    
    results = {}
    for filepath in workbooks:
        success = validate_workbook(filepath)
        results[os.path.basename(filepath)] = success
        print()
    
    # Final Summary
    print("=" * 70)
    print("BATCH SUMMARY")
    print("=" * 70)
    
    passed = sum(1 for v in results.values() if v)
    failed = len(results) - passed
    
    for filename, success in results.items():
        status = "[PASS]" if success else "[FAIL]"
        print(f"  {status}  {filename}")
    
    print(f"\nTotal: {passed} passed, {failed} failed out of {len(results)}")
    print("=" * 70)
    
    return failed == 0


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def print_usage():
    """Print usage information."""
    print("""
Excel Sanity Check - ISMS Control 5.23 Cloud Services Workbooks
========================================================================

Usage:
    python3 excel_sanity_check_a523.py <workbook.xlsx>
    python3 excel_sanity_check_a523.py --all
    python3 excel_sanity_check_a523.py --help

Supported Workbooks:
    5.23.1 - Cloud Service Inventory      (ISMS-IMP-A.5.23.S1_Inventory_*.xlsx)
    5.23.2 - Vendor Due Diligence         (ISMS-IMP-A.5.23.S2_VendorDD_*.xlsx)
    5.23.3 - Secure Configuration         (ISMS-IMP-A.5.23.S3_SecureConfig_*.xlsx)
    5.23.4 - Ongoing Governance           (ISMS-IMP-A.5.23.S4_Governance_*.xlsx)
    5.23.5 - Compliance & Exit Planning   (ISMS-IMP-A.5.23.S5_Compliance_*.xlsx)

Validation Checks:
    [OK] Sheet structure (all expected sheets present)
    [OK] Column headers (base columns match spec)
    [OK] Data validations (dropdowns present)
    [OK] Dashboard formulas (auto-calculation)
    [OK] Evidence Register format
    [OK] Approval Sign-Off structure
    [OK] Checklist sections (5.23.3-5.23.5)

Examples:
    python3 excel_sanity_check_a523.py ISMS-IMP-A.5.23.S4_Governance_20260101.xlsx
    python3 excel_sanity_check_a523.py --all
""")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print_usage()
        sys.exit(1)
    
    arg = sys.argv[1]
    
    if arg in ["--help", "-h", "help"]:
        print_usage()
        sys.exit(0)
    
    if arg == "--all":
        success = validate_all_workbooks()
        sys.exit(0 if success else 1)
    
    if not os.path.exists(arg):
        print(f"[X] ERROR: File not found: {arg}")
        sys.exit(1)
    
    success = validate_workbook(arg)
    sys.exit(0 if success else 1)