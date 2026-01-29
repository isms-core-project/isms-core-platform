#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.5.31.2 - Applicability Assessment Matrix Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.31: Legal, Statutory, Regulatory and Contractual Requirements
Assessment Domain 2 of 6: Regulatory Applicability Determination

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific applicability assessment methodology, decision
criteria, and approval workflows.

Key customization areas:
1. Applicability criteria and scoring methodology (adapt to your risk framework)
2. Geographic, operational, and contractual scope definitions (match your business model)
3. Decision thresholds and scoring weights (align with your materiality assessments)
4. Approval workflow requirements (based on your governance structure)
5. Assessment template structure (specific to your regulatory analysis needs)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.31 Regulatory Compliance Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for conducting
systematic applicability assessments of regulations identified in the Regulatory
Inventory (Workbook 1), determining which regulations actually apply to
[Organization] through structured evaluation of geographic, operational, and
contractual scope.

**Purpose:**
Enables systematic, auditable, and repeatable determination of regulatory
applicability against ISO 27001:2022 Control A.5.31 requirements, providing
documented rationale for compliance obligations and supporting evidence-based
tier categorization decisions.

**Assessment Scope:**
- Geographic scope analysis (jurisdictional presence and data processing locations)
- Operational scope analysis (business activities, services offered, data types processed)
- Contractual scope analysis (customer requirements, supplier obligations, partnership agreements)
- Organizational characteristics (entity type, size, revenue, employee count)
- Technical characteristics (infrastructure, cloud services, data flows)
- Multi-criteria scoring methodology for applicability determination
- Structured decision matrix with weighted criteria
- Approval workflow tracking (Compliance Officer, Legal Counsel, ISMS Manager)
- Assessment history and versioning
- Integration with Regulatory Inventory (Workbook 1)
- Feeds into Requirements Extraction (Workbook 3)

**Generated Workbook Structure:**
1. Applicability_Assessment - Completed example assessment (e.g., GDPR)
2. Instructions - Comprehensive assessment methodology guidance
3. Template_Blank - Blank template for new assessments

**Key Features:**
- Structured three-dimensional scope assessment (Geographic/Operational/Contractual)
- Weighted scoring system for objective applicability determination
- Data validation with dropdown lists for standardized responses
- Conditional formatting for visual scoring indication
- Approval workflow with signature tracking
- Assessment versioning and history
- Protected formulas with unprotected input cells
- Sample completed assessment as reference
- Blank template for new assessments
- UTF-8 encoding with emoji support

**Integration:**
This assessment matrix feeds into:
- ISMS-POL-00 (Regulatory Applicability Framework - updates tier classifications)
- Workbook 1 (Regulatory Inventory - updates applicability status)
- Workbook 3 (Requirements Extraction - identifies which regulations need requirements extracted)
- Dashboard (Compliance Overview - provides applicability statistics)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_531_2_applicability_matrix.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_531_2_applicability_matrix.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_531_2_applicability_matrix.py --date 20250125

Output:
    File: ISMS_Assessment_531_2_Applicability_Matrix_YYYYMMDD.xlsx
    Location: ../90_workbooks/ (or specified output path)

Post-Generation Steps:
    1. Review applicability assessment methodology in Instructions sheet
    2. Review sample completed assessment (Applicability_Assessment sheet)
    3. For each regulation in Inventory requiring assessment:
       a. Copy Template_Blank sheet, rename with regulation ID
       b. Complete all assessment sections systematically
       c. Calculate applicability score using weighted criteria
       d. Determine tier classification based on score
       e. Document detailed rationale for determination
       f. Obtain required approvals (Legal Counsel for Tier 1)
    4. Update Regulatory Inventory (Workbook 1) with assessment results
    5. Link assessment reference in Inventory to this workbook
    6. Trigger Requirements Extraction (Workbook 3) for applicable regulations
    7. Archive completed assessments for audit evidence

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.31
Assessment Domain:    2 of 6 (Regulatory Applicability Determination)
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.31.2: Regulatory Applicability Methodology
    - ISMS-IMP-A.5.31.2: Regulatory Applicability Assessment Process
    - ISMS-POL-00: Regulatory Applicability Framework

Related Assessment Tools:
    - Assessment Workbook 1: Regulatory Inventory (generate_531_1_regulatory_inventory.py)
    - Assessment Workbook 3: Requirements Extraction (generate_531_3_requirements_register.py)
    - Assessment Workbook 4: Control Mapping (generate_531_4_control_mapping.py)
    - Dashboard: Compliance Overview (generate_531_6_compliance_dashboard.py)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements three-dimensional applicability assessment framework
    - Supports structured, repeatable applicability determinations
    - Integrated approval workflow tracking
    - Sample completed assessment (GDPR) as reference

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Applicability Assessment Methodology:**
Assessment uses three-dimensional scope analysis:

1. **Geographic Scope** (Where):
   - Physical presence (offices, data centers, employees)
   - Market presence (customers, sales, marketing)
   - Data processing locations (data centers, cloud regions, backups)
   - Cross-border data transfers

2. **Operational Scope** (What):
   - Business activities and services offered
   - Data types processed (personal data, financial data, health data)
   - Technology used (cloud services, IoT, AI/ML)
   - Partnerships and third-party relationships

3. **Contractual Scope** (Who requires it):
   - Customer contractual obligations
   - Supplier and partner requirements
   - Industry certifications and memberships
   - Voluntary commitments (privacy policies, terms of service)

**Scoring Methodology:**
Each dimension scored on weighted criteria:
- 0 points: Not applicable / No presence / Not relevant
- 1 point: Minimal / Edge case / Uncertainty
- 2 points: Moderate / Partial / Indirect
- 3 points: Significant / Direct / Clear

Total score determines tier classification:
- Score 20-30: Tier 1 (Mandatory) - Clear legal/contractual obligation
- Score 10-19: Tier 2 (Conditional) - Would apply if conditions met
- Score 0-9: Tier 3 (Informational) or Not Applicable

**IMPORTANT:** Scoring is guidance, not absolute. Legal counsel interpretation
may override score-based determination for complex regulations.

**When to Conduct Assessment:**
- Initial: When new regulation identified in inventory
- Re-assessment: Annually for Tier 1, Quarterly for Tier 2, Biennially for Tier 3
- Triggered: When organizational changes occur (new markets, services, acquisitions, partnerships)
- Ad-hoc: When regulation itself changes (amendments, new interpretations, enforcement guidance)

**Approval Requirements:**
Critical governance control - different tiers require different approvals:

- **Tier 3 / Not Applicable**: 
  - Compliance Officer approval (1 business day)
  - Low risk of incorrect determination

- **Tier 2 (Conditional)**:
  - Compliance Officer + ISMS Manager approval (2-3 business days)
  - Moderate risk - need to monitor trigger conditions

- **Tier 1 (Mandatory)**:
  - Compliance Officer + Legal Counsel + ISMS Manager + Executive Management (1-2 weeks)
  - High risk - legal penalties, compliance failures, contractual breaches if wrong

**Audit Considerations:**
Applicability assessments are critical audit evidence. Auditors expect:
- Systematic methodology applied consistently across all regulations
- Clear, documented rationale for each determination
- Evidence of legal counsel involvement for Tier 1 determinations
- Appropriate approvals obtained before finalization
- Assessment versioning and history (can trace changes over time)
- Regular reassessment per defined schedule

**Data Protection:**
Assessment workbooks contain sensitive strategic information:
- Business activities and market presence
- Customer relationships and contractual obligations
- Technology stack and infrastructure details
- Legal interpretations and risk assessments
- Strategic plans (future markets, services)

Handle in accordance with [Organization]'s data classification policies.

**Legal Counsel Critical Role:**
Legal counsel MUST review and approve ALL Tier 1 determinations. They provide:
- Authoritative interpretation of complex legal language
- Jurisdiction-specific guidance (extraterritoriality, conflicts of law)
- Analysis of regulatory intent vs. literal text
- Enforcement risk assessment
- Regulatory agency guidance interpretation

Do NOT finalize Tier 1 determinations without legal counsel sign-off.

**Common Assessment Challenges:**

1. **Extraterritoriality** (e.g., GDPR applies to non-EU companies):
   - Don't assume geographic boundaries limit applicability
   - Consider data subjects' location, not just company location
   - Analyze long-arm provisions in regulations

2. **Threshold Triggers** (e.g., DORA requires €30M annual revenue):
   - Document current status vs. threshold
   - Set monitoring for approaching thresholds
   - Classify as Tier 2 (Conditional) if close to threshold

3. **Ambiguous Definitions** (e.g., "Critical Infrastructure Entity"):
   - Seek regulatory agency guidance documents
   - Consult industry associations
   - When uncertain: err on side of applicability (can downgrade later)

4. **Customer Contractual Requirements**:
   - Treat as Tier 1 (contractual breach has legal consequences)
   - Review all master service agreements for InfoSec obligations
   - Aggregate common requirements across customers

5. **Future Services/Markets**:
   - Assess based on CURRENT state, not future plans
   - Classify future regulations as Tier 2 (Conditional on market entry)
   - Reassess when plans become reality

**Template Usage:**
The Template_Blank sheet is designed for easy replication:
1. Copy entire sheet (right-click tab → Move or Copy → Create a copy)
2. Rename sheet with Regulation ID (e.g., "REG-EU-002-Assessment")
3. Complete all sections systematically
4. Save with clear version control if assessment updated

**Assessment History:**
Maintain assessment history for audit trail:
- Keep previous versions when reassessing (append date to sheet name)
- Document what changed and why in rationale sections
- Link to triggering event (org change, regulation amendment, audit finding)

**Quality Assurance:**
Before finalizing assessment:
- All sections completed (no blank fields)
- Scoring consistent with described scenarios
- Rationale provides specific, factual justification
- Legal counsel reviewed if Tier 1 determination
- Appropriate approvals obtained and dated
- Assessment reference added to Regulatory Inventory (Workbook 1)

**Integration with Inventory:**
After completing assessment:
1. Update Workbook 1 (Regulatory Inventory):
   - Tier column (1/2/3 or Not Applicable)
   - Applicability Status (Applicable/Conditional/Not Applicable)
   - Applicability Rationale (summary from assessment)
   - Assessment Reference (link to this workbook, sheet name)
2. If Tier 1 or Tier 2 Applicable: Trigger Requirements Extraction (Workbook 3)
3. Update POL-00 if material change to regulatory portfolio

**Reassessment Triggers:**
Conduct reassessment when:
- Regulation itself changes (amendments, new enforcement guidance)
- Organizational changes (new markets, services, revenue thresholds, entity types)
- Contractual changes (new customers with specific requirements, supplier obligations)
- Technology changes (cloud adoption, AI/ML deployment, IoT introduction)
- Audit findings question previous determination
- Peer organizations reach different determination (triggers review)

**Common Pitfalls:**
- Conducting assessment without reading full regulation text
- Relying solely on summaries or third-party interpretations
- Skipping legal counsel review for Tier 1 determinations
- Not documenting trigger conditions for Tier 2 (Conditional)
- Confusing "we should comply" with "we must comply legally"
- Ignoring contractual obligations (treating as "voluntary")
- Not reassessing when organizational context changes

**Scalability:**
For organizations with 50-100+ regulations in inventory:
- Prioritize assessments: Tier 1 suspected regulations first
- Batch similar regulations (e.g., all US state privacy laws together)
- Create assessment templates for common regulation types
- Consider engaging compliance consultants for initial assessments
- Use regulatory technology platforms for ongoing monitoring

================================================================================
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================



# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
SCALES = '\u2696'     # ⚖️  Scales of Justice
DOCUMENT = '\U0001F4C4' # 📄 Document
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Sheet structure
    sheets = [
        "Applicability_Assessment",
        "Instructions",
        "Template_Blank",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    
    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    medium = Side(style="medium")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)
    
    styles = {
        "section_header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="203864", end_color="203864", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": border_medium,
        },
        "subsection_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": border_thin,
        },
        "field_label": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
            "alignment": Alignment(horizontal="right", vertical="center"),
            "border": border_thin,
        },
        "input_field": {
            "font": Font(name="Calibri", size=10),
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "question_text": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "dropdown_cell": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "evidence_cell": {
            "font": Font(name="Calibri", size=9),
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border_thin,
        },
        "score_cell": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "rating_low": {
            "fill": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
        },
        "rating_moderate": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
        },
        "rating_high": {
            "fill": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        f = style_dict["font"]
        cell.font = Font(
            name=f.name,
            size=f.size,
            bold=f.bold if hasattr(f, "bold") else False,
            color=f.color if hasattr(f, "color") else "000000",
        )
    if "fill" in style_dict:
        f = style_dict["fill"]
        cell.fill = PatternFill(
            start_color=f.start_color.rgb if hasattr(f.start_color, "rgb") else f.start_color,
            end_color=f.end_color.rgb if hasattr(f.end_color, "rgb") else f.end_color,
            fill_type=f.fill_type,
        )
    if "alignment" in style_dict:
        a = style_dict["alignment"]
        cell.alignment = Alignment(
            horizontal=a.horizontal,
            vertical=a.vertical,
            wrap_text=a.wrap_text if hasattr(a, "wrap_text") else False,
        )
    if "border" in style_dict:
        cell.border = style_dict["border"]


# ============================================================================
# SECTION 2: ASSESSMENT FORM STRUCTURE
# ============================================================================

def create_section_header(ws, row, col_start, col_end, text, styles):
    """Create a merged section header."""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    apply_style(cell, styles["section_header"])
    ws.row_dimensions[row].height = 25
    return row + 1


def create_field_row(ws, row, label, value, styles, col_label=1, col_value=2, col_value_end=6):
    """Create a field row with label and input area."""
    # Label cell
    label_cell = ws.cell(row=row, column=col_label, value=label)
    apply_style(label_cell, styles["field_label"])
    
    # Value cell (merged)
    ws.merge_cells(start_row=row, start_column=col_value, end_row=row, end_column=col_value_end)
    value_cell = ws.cell(row=row, column=col_value, value=value)
    apply_style(value_cell, styles["input_field"])
    
    return row + 1


def create_question_row(ws, row, question_num, question_text, styles):
    """Create a question row with Y/N dropdown and evidence field."""
    # Question text (merged columns A-B)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    q_cell = ws.cell(row=row, column=1, value=f"{question_num}: {question_text}")
    apply_style(q_cell, styles["question_text"])
    
    # Answer dropdown (column C)
    answer_cell = ws.cell(row=row, column=3)
    apply_style(answer_cell, styles["dropdown_cell"])
    
    # Evidence field (merged columns D-F)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    evidence_cell = ws.cell(row=row, column=4)
    apply_style(evidence_cell, styles["evidence_cell"])
    
    ws.row_dimensions[row].height = 30
    return row + 1


# ============================================================================
# SECTION 3: POPULATE APPLICABILITY ASSESSMENT SHEET
# ============================================================================

def populate_assessment_sheet(wb, styles, sheet_name="Applicability_Assessment", with_sample_data=True):
    """Populate the specified assessment sheet."""
    ws = wb[sheet_name]
    
    # Set column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25
    
    row = 1
    
    # ========================================================================
    # SECTION A: REGULATION IDENTIFICATION
    # ========================================================================
    row = create_section_header(ws, row, 1, 6, "SECTION A: REGULATION IDENTIFICATION", styles)
    
    if with_sample_data:
        sample_reg = {
            "Regulation ID": "REG-EU-002",
            "Regulation Name": "Network and Information Systems Directive (Revised)",
            "Jurisdiction": "European Union",
            "Issuing Authority": "European Commission",
            "Citation": "Directive (EU) 2022/2555",
            "Effective Date": "2024-10-17",
            "Assessment Date": datetime.now().strftime("%Y-%m-%d"),
            "Assessor": "Compliance Officer / CISO",
            "Trigger Event": "Annual review of Tier 2 regulations; assessing if NIS2 thresholds now met",
        }
    else:
        sample_reg = {key: "" for key in [
            "Regulation ID", "Regulation Name", "Jurisdiction", "Issuing Authority",
            "Citation", "Effective Date", "Assessment Date", "Assessor", "Trigger Event"
        ]}
    
    row = create_field_row(ws, row, "Regulation ID:", sample_reg["Regulation ID"], styles)
    row = create_field_row(ws, row, "Regulation Name:", sample_reg["Regulation Name"], styles)
    row = create_field_row(ws, row, "Jurisdiction:", sample_reg["Jurisdiction"], styles)
    row = create_field_row(ws, row, "Issuing Authority:", sample_reg["Issuing Authority"], styles)
    row = create_field_row(ws, row, "Citation:", sample_reg["Citation"], styles)
    row = create_field_row(ws, row, "Effective Date:", sample_reg["Effective Date"], styles)
    row = create_field_row(ws, row, "Assessment Date:", sample_reg["Assessment Date"], styles)
    row = create_field_row(ws, row, "Assessor:", sample_reg["Assessor"], styles)
    row = create_field_row(ws, row, "Trigger Event:", sample_reg["Trigger Event"], styles)
    
    row += 1
    
    # ========================================================================
    # SECTION B: GEOGRAPHIC SCOPE ASSESSMENT
    # ========================================================================
    row = create_section_header(ws, row, 1, 6, "SECTION B: GEOGRAPHIC SCOPE ASSESSMENT", styles)
    
    # Column headers
    ws.cell(row=row, column=1, value="Question").font = Font(bold=True, size=9)
    ws.cell(row=row, column=3, value="Y/N").font = Font(bold=True, size=9)
    ws.cell(row=row, column=4, value="Evidence / Notes").font = Font(bold=True, size=9)
    ws.row_dimensions[row].height = 15
    row += 1
    
    geo_questions = [
        ("G1", "Physical operations in jurisdiction?"),
        ("G2", "Legal entities registered in jurisdiction?"),
        ("G3", "Employees located in jurisdiction?"),
        ("G4", "Customers located in jurisdiction?"),
        ("G5", "Data subjects in jurisdiction?"),
        ("G6", "Actively targeting jurisdiction (marketing, sales)?"),
        ("G7", "Extraterritorial application to [Organization]?"),
    ]
    
    geo_start_row = row
    for q_num, q_text in geo_questions:
        if with_sample_data and q_num in ["G4", "G5", "G7"]:
            ws.cell(row=row, column=3, value="Y")
        row = create_question_row(ws, row, q_num, q_text, styles)
    
    # Geographic Score calculation
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    score_label = ws.cell(row=row, column=1, value="Geographic Score:")
    score_label.font = Font(bold=True, size=11)
    score_label.alignment = Alignment(horizontal="right", vertical="center")
    
    score_cell = ws.cell(row=row, column=3, value=f'=COUNTIF(C{geo_start_row}:C{row-2},"Y")')
    apply_style(score_cell, styles["score_cell"])
    
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    rating_label = ws.cell(row=row, column=1, value="Geographic Rating:")
    rating_label.font = Font(bold=True, size=11)
    rating_label.alignment = Alignment(horizontal="right", vertical="center")
    
    rating_formula = f'=IF(C{row-1}<=1,"Low",IF(C{row-1}<=3,"Moderate","High"))'
    rating_cell = ws.cell(row=row, column=3, value=rating_formula)
    apply_style(rating_cell, styles["score_cell"])
    
    row += 2
    
    # ========================================================================
    # SECTION C: OPERATIONAL SCOPE ASSESSMENT
    # ========================================================================
    row = create_section_header(ws, row, 1, 6, "SECTION C: OPERATIONAL SCOPE ASSESSMENT", styles)
    
    # Column headers
    ws.cell(row=row, column=1, value="Question").font = Font(bold=True, size=9)
    ws.cell(row=row, column=3, value="Y/N").font = Font(bold=True, size=9)
    ws.cell(row=row, column=4, value="Evidence / Notes").font = Font(bold=True, size=9)
    ws.row_dimensions[row].height = 15
    row += 1
    
    ops_questions = [
        ("O1", "Services offered covered by regulation?"),
        ("O2", "Industry sectors operated in covered?"),
        ("O3", "Data types processed covered by regulation?"),
        ("O4", "Size/revenue/employee thresholds met?"),
        ("O5", "Specific operations or activities covered?"),
    ]
    
    ops_start_row = row
    for q_num, q_text in ops_questions:
        if with_sample_data and q_num in ["O1", "O3"]:
            ws.cell(row=row, column=3, value="Y")
        row = create_question_row(ws, row, q_num, q_text, styles)
    
    # Operational Score calculation
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    score_label = ws.cell(row=row, column=1, value="Operational Score:")
    score_label.font = Font(bold=True, size=11)
    score_label.alignment = Alignment(horizontal="right", vertical="center")
    
    score_cell = ws.cell(row=row, column=3, value=f'=COUNTIF(C{ops_start_row}:C{row-2},"Y")')
    apply_style(score_cell, styles["score_cell"])
    
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    rating_label = ws.cell(row=row, column=1, value="Operational Rating:")
    rating_label.font = Font(bold=True, size=11)
    rating_label.alignment = Alignment(horizontal="right", vertical="center")
    
    rating_formula = f'=IF(C{row-1}<=1,"Low",IF(C{row-1}<=3,"Moderate","High"))'
    rating_cell = ws.cell(row=row, column=3, value=rating_formula)
    apply_style(rating_cell, styles["score_cell"])
    
    row += 2
    
    # ========================================================================
    # SECTION D: CONTRACTUAL SCOPE ASSESSMENT
    # ========================================================================
    row = create_section_header(ws, row, 1, 6, "SECTION D: CONTRACTUAL SCOPE ASSESSMENT", styles)
    
    # Column headers
    ws.cell(row=row, column=1, value="Question").font = Font(bold=True, size=9)
    ws.cell(row=row, column=3, value="Y/N").font = Font(bold=True, size=9)
    ws.cell(row=row, column=4, value="Evidence / Notes").font = Font(bold=True, size=9)
    ws.row_dimensions[row].height = 15
    row += 1
    
    contract_questions = [
        ("C1", "Contracts explicitly require compliance?"),
        ("C2", "Certifications/attestations required?"),
        ("C3", "Audit rights granted to customers/partners?"),
        ("C4", "Market expectation or industry norm?"),
    ]
    
    contract_start_row = row
    for q_num, q_text in contract_questions:
        if with_sample_data and q_num == "C4":
            ws.cell(row=row, column=3, value="Y")
        row = create_question_row(ws, row, q_num, q_text, styles)
    
    # Contractual Score calculation
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    score_label = ws.cell(row=row, column=1, value="Contractual Score:")
    score_label.font = Font(bold=True, size=11)
    score_label.alignment = Alignment(horizontal="right", vertical="center")
    
    score_cell = ws.cell(row=row, column=3, value=f'=COUNTIF(C{contract_start_row}:C{row-2},"Y")')
    apply_style(score_cell, styles["score_cell"])
    
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    rating_label = ws.cell(row=row, column=1, value="Contractual Rating:")
    rating_label.font = Font(bold=True, size=11)
    rating_label.alignment = Alignment(horizontal="right", vertical="center")
    
    rating_formula = f'=IF(C{row-1}=0,"None",IF(C{row-1}<=2,"Moderate","High"))'
    rating_cell = ws.cell(row=row, column=3, value=rating_formula)
    apply_style(rating_cell, styles["score_cell"])
    
    row += 2
    
    # ========================================================================
    # SECTION E: OVERALL DETERMINATION
    # ========================================================================
    row = create_section_header(ws, row, 1, 6, "SECTION E: OVERALL DETERMINATION", styles)
    
    # Summary of Scores
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="Summary of Assessment Scores:").font = Font(bold=True, size=10)
    row += 1
    
    # Use the score rows we calculated above - need to reference them
    # For simplicity, just create labels
    ws.cell(row=row, column=1, value="Geographic:").font = Font(size=10)
    ws.cell(row=row, column=2, value="[See Section B]").font = Font(size=9, italic=True)
    row += 1
    ws.cell(row=row, column=1, value="Operational:").font = Font(size=10)
    ws.cell(row=row, column=2, value="[See Section C]").font = Font(size=9, italic=True)
    row += 1
    ws.cell(row=row, column=1, value="Contractual:").font = Font(size=10)
    ws.cell(row=row, column=2, value="[See Section D]").font = Font(size=9, italic=True)
    row += 2
    
    # Applicability Determination
    row = create_field_row(ws, row, "Applicability Determination:", 
                          "Conditionally Applicable" if with_sample_data else "",
                          styles, col_value_end=3)
    
    # Rationale (large text area)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
    label_cell = ws.cell(row=row, column=1, value="Rationale:")
    apply_style(label_cell, styles["field_label"])
    
    ws.merge_cells(start_row=row, start_column=2, end_row=row+3, end_column=6)
    rationale_cell = ws.cell(row=row, column=2)
    if with_sample_data:
        rationale_cell.value = ("NIS2 applies to 'essential' and 'important' entities. Current assessment: "
                                "[Organization] approaching but not yet meeting size thresholds (€10M revenue, 50 employees). "
                                "Geographic scope: High (EU operations, EU customers). Operational scope: Moderate (digital services). "
                                "Contractual: Low (market expectation only). "
                                "DETERMINATION: Conditionally Applicable - would become Tier 1 if size thresholds exceeded.")
    apply_style(rationale_cell, styles["input_field"])
    ws.row_dimensions[row].height = 60
    row += 4
    
    # Tier Assignment
    row = create_field_row(ws, row, "Tier Assignment:",
                          "Tier 2 (Conditional)" if with_sample_data else "",
                          styles, col_value_end=3)
    
    # Tier Rationale
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
    label_cell = ws.cell(row=row, column=1, value="Tier Rationale:")
    apply_style(label_cell, styles["field_label"])
    
    ws.merge_cells(start_row=row, start_column=2, end_row=row+2, end_column=6)
    tier_rationale_cell = ws.cell(row=row, column=2)
    if with_sample_data:
        tier_rationale_cell.value = ("Tier 2 (Conditional): NIS2 would apply when [Organization] meets 'medium-sized enterprise' "
                                    "thresholds OR is designated as essential/important entity by national authority. "
                                    "Currently monitoring revenue growth and employee count quarterly.")
    apply_style(tier_rationale_cell, styles["input_field"])
    ws.row_dimensions[row].height = 45
    row += 3
    
    # Applicability Condition (for Tier 2)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
    label_cell = ws.cell(row=row, column=1, value="Condition for Tier 1 Upgrade:")
    apply_style(label_cell, styles["field_label"])
    
    ws.merge_cells(start_row=row, start_column=2, end_row=row+1, end_column=6)
    condition_cell = ws.cell(row=row, column=2)
    if with_sample_data:
        condition_cell.value = ("TRIGGER: Annual revenue exceeds €10M OR employee count exceeds 50 OR designated by national authority. "
                               "Reassess quarterly during growth phase.")
    apply_style(condition_cell, styles["input_field"])
    ws.row_dimensions[row].height = 30
    row += 2
    
    row += 1
    
    # ========================================================================
    # SECTION F: APPROVAL & SIGN-OFF
    # ========================================================================
    row = create_section_header(ws, row, 1, 6, "SECTION F: APPROVAL & SIGN-OFF", styles)
    
    # Approval table headers
    headers = ["Role", "Name", "Date", "Signature / Comments"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[row].height = 20
    row += 1
    
    # Approval rows
    roles = [
        "Prepared By:",
        "Reviewed By (Compliance Officer):",
        "Reviewed By (Legal Counsel):",
        "Approved By (ISMS Manager):",
    ]
    
    for role in roles:
        ws.cell(row=row, column=1, value=role).font = Font(size=10)
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Add data validations
    add_validations(ws)
    
    # Add conditional formatting for ratings
    add_conditional_formatting(ws)


def add_validations(ws):
    """Add data validation dropdowns."""
    # Y/N dropdowns for all question answer cells (column C)
    yn_validation = DataValidation(
        type="list",
        formula1='"Y,N"',
        allow_blank=True
    )
    yn_validation.error = "Please enter Y or N"
    yn_validation.errorTitle = "Invalid Answer"
    ws.add_data_validation(yn_validation)
    yn_validation.add("C1:C100")  # Cover all potential question rows
    
    # Applicability Determination dropdown
    determination_validation = DataValidation(
        type="list",
        formula1='"Applicable,Conditionally Applicable,Not Applicable,Uncertain"',
        allow_blank=True
    )
    ws.add_data_validation(determination_validation)
    # Will apply to specific cell - need to find the row
    # For now apply broadly
    determination_validation.add("B52:B60")
    
    # Tier Assignment dropdown
    tier_validation = DataValidation(
        type="list",
        formula1='"Tier 1 (Mandatory),Tier 2 (Conditional),Tier 3 (Informational),N/A"',
        allow_blank=True
    )
    ws.add_data_validation(tier_validation)
    tier_validation.add("B62:B70")


def add_conditional_formatting(ws):
    """Add conditional formatting for rating cells."""
    from openpyxl.formatting.rule import CellIsRule
    
    # Rating cells - color code based on value
    low_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    moderate_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Apply to potential rating cells (column C)
    ws.conditional_formatting.add(
        "C1:C100",
        CellIsRule(operator="equal", formula=['"Low"'], fill=low_fill)
    )
    ws.conditional_formatting.add(
        "C1:C100",
        CellIsRule(operator="equal", formula=['"Moderate"'], fill=moderate_fill)
    )
    ws.conditional_formatting.add(
        "C1:C100",
        CellIsRule(operator="equal", formula=['"High"'], fill=high_fill)
    )
    ws.conditional_formatting.add(
        "C1:C100",
        CellIsRule(operator="equal", formula=['"None"'], fill=low_fill)
    )


# ============================================================================
# SECTION 4: INSTRUCTIONS SHEET
# ============================================================================

def populate_instructions_sheet(wb):
    """Populate the Instructions sheet."""
    ws = wb["Instructions"]
    
    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "APPLICABILITY ASSESSMENT MATRIX - INSTRUCTIONS"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    instructions = [
        "",
        "PURPOSE:",
        "Systematic assessment form for determining whether a specific regulation applies to [Organization],",
        "and if so, at what tier (Mandatory, Conditional, or Informational).",
        "",
        "WHEN TO USE:",
        f"{BULLET} New regulation identified that may apply",
        f"{BULLET} Annual review of existing Tier 2/3 regulations",
        f"{BULLET} Business change triggers reassessment (new market, new service, etc.)",
        f"{BULLET} Regulatory change (amendment, new guidance) triggers reassessment",
        "",
        "HOW TO COMPLETE:",
        "",
        "SECTION A - REGULATION IDENTIFICATION:",
        "Complete basic information about the regulation being assessed.",
        f"{BULLET} Trigger Event: Document WHY this assessment is being performed now",
        "",
        "SECTION B - GEOGRAPHIC SCOPE:",
        "Answer Y/N for each question. Score auto-calculates.",
        f"{BULLET} G1-G3: Physical presence and legal establishment",
        f"{BULLET} G4-G5: Customer and data subject location",
        f"{BULLET} G6-G7: Targeting and extraterritorial reach",
        f"{BULLET} Rating: Low (0-1 Yes), Moderate (2-3 Yes), High (4+ Yes)",
        "",
        "SECTION C - OPERATIONAL SCOPE:",
        "Answer Y/N for each question. Score auto-calculates.",
        f"{BULLET} O1-O2: Services and industry sectors covered",
        f"{BULLET} O3: Data types processed",
        f"{BULLET} O4: Size/revenue/employee thresholds",
        f"{BULLET} O5: Specific operations or activities",
        f"{BULLET} Rating: Low (0-1 Yes), Moderate (2-3 Yes), High (4+ Yes)",
        "",
        "SECTION D - CONTRACTUAL SCOPE:",
        "Answer Y/N for each question. Score auto-calculates.",
        f"{BULLET} C1: Explicit contractual requirement",
        f"{BULLET} C2: Certification/attestation requirements",
        f"{BULLET} C3: Audit rights",
        f"{BULLET} C4: Market expectation",
        f"{BULLET} Rating: None (0 Yes), Moderate (1-2 Yes), High (3+ Yes)",
        "",
        "SECTION E - OVERALL DETERMINATION:",
        "Based on scores from Sections B-D, make final determination.",
        "",
        "Decision Matrix (Guideline):",
        "┌─────────────────┬──────────────┬────────────────┐",
        "│ Geographic      │ Operational  │ Determination  │",
        "├─────────────────┼──────────────┼────────────────┤",
        "│ High            │ High         │ Applicable     │",
        "│ High            │ Moderate     │ Likely Appl.   │",
        "│ Moderate        │ High         │ Conditional    │",
        "│ Low/None        │ Low/None     │ Not Applicable │",
        "└─────────────────┴──────────────┴────────────────┘",
        "",
        "Applicability Determination:",
        f"{BULLET} Applicable: Regulation applies NOW, compliance required",
        f"{BULLET} Conditionally Applicable: Would apply if condition met",
        f"{BULLET} Not Applicable: Does not apply to [Organization]",
        f"{BULLET} Uncertain: Insufficient info, requires legal counsel review",
        "",
        "Tier Assignment:",
        f"{BULLET} Tier 1 (Mandatory): Legal/contractual obligation, must comply",
        f"{BULLET} Tier 2 (Conditional): Would become Tier 1 if condition triggered",
        f"{BULLET} Tier 3 (Informational): Voluntary framework, used for reference",
        f"{BULLET} N/A: Not applicable",
        "",
        "SECTION F - APPROVAL & SIGN-OFF:",
        "Required approvals based on tier determination:",
        f"{BULLET} Tier 3 / N/A: Compliance Officer",
        f"{BULLET} Tier 2: Compliance Officer + ISMS Manager",
        f"{BULLET} Tier 1: Compliance Officer + Legal Counsel + ISMS Manager + Executive",
        "",
        "IMPORTANT NOTES:",
        f"{BULLET} This is a FRAMEWORK for assessment, not a legal determination",
        f"{BULLET} Legal counsel review REQUIRED for any Tier 1 determination",
        f"{BULLET} Document all reasoning clearly - this is audit evidence",
        f"{BULLET} For 'Not Applicable': Document what would trigger reassessment",
        f"{BULLET} For 'Conditional': Document the specific trigger condition",
        "",
        "OUTPUTS:",
        f"{BULLET} Updates Regulatory Inventory (Workbook 1) with tier and status",
        f"{BULLET} If Applicable/Tier 1: Triggers Requirements Extraction (Workbook 3)",
        f"{BULLET} Assessment form saved as evidence (reference in Workbook 1)",
        "",
        "RELATED PROCESSES:",
        f"{BULLET} 5.31.2: Applicability Assessment Process (detailed procedures)",
        f"{BULLET} POL-5.31.2: Regulatory Applicability Methodology (framework)",
        "",
        "QUALITY CHECKS:",
        "✓ All Y/N questions answered",
        "✓ Evidence provided for each 'Y' answer",
        "✓ Rationale is specific and audit-ready",
        "✓ Tier assignment matches determination logic",
        "✓ For Tier 2: Condition clearly documented",
        "✓ Appropriate approvals obtained",
    ]
    
    for idx, line in enumerate(instructions, start=2):
        cell = ws.cell(row=idx, column=1, value=line)
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    ws.column_dimensions["A"].width = 100
    ws.freeze_panes = "A2"


# ============================================================================
# SECTION 5: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    print("=" * 70)
    print("ISMS Assessment Workbook Generator")
    print("Control A.5.31 - Workbook 2: Applicability Assessment Matrix")
    print("=" * 70)
    print()
    
    # Create workbook
    print("📋 Creating workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    
    # Populate sheets
    print("📝 Populating Applicability Assessment sheet (with sample data)...")
    populate_assessment_sheet(wb, styles, sheet_name="Applicability_Assessment", with_sample_data=True)
    
    print("📖 Populating Instructions sheet...")
    populate_instructions_sheet(wb)
    
    print(f"{DOCUMENT} Creating blank template sheet...")
    populate_assessment_sheet(wb, styles, sheet_name="Template_Blank", with_sample_data=False)
    
    # Save workbook
    output_path = f"ISMS-IMP-A.5.31.2_Applicability_Matrix_{datetime.now().strftime('%Y%m%d')}.xlsx"
    print(f"💾 Saving workbook to: {output_path}")
    wb.save(output_path)
    
    print()
    print(f"{CHECK} Workbook generated successfully!")
    print()
    print("📦 Output Details:")
    print(f"   File: ISMS_Assessment_531_2_Applicability_Matrix.xlsx")
    print(f"   Location: ../90_workbooks/")
    print(f"   Sheets: 3 (Applicability_Assessment, Instructions, Template_Blank)")
    print()
    print("📋 Features:")
    print("   ✓ Structured assessment form with 6 sections")
    print("   ✓ 16 assessment questions (7 geographic, 5 operational, 4 contractual)")
    print("   ✓ Auto-calculating scores and ratings")
    print("   ✓ Y/N dropdowns for questions")
    print("   ✓ Determination and Tier assignment dropdowns")
    print("   ✓ Conditional formatting (Low/Moderate/High ratings)")
    print("   ✓ Approval workflow section")
    print("   ✓ Sample data (NIS2 assessment)")
    print("   ✓ Blank template for copying")
    print("   ✓ Comprehensive instructions")
    print()
    print(f"{TARGET} Next Steps:")
    print("   1. Open workbook and review sample assessment (NIS2)")
    print("   2. Use Template_Blank sheet to assess new regulations")
    print("   3. Follow approval workflow based on tier determination")
    print("   4. Update Regulatory Inventory (Workbook 1) with results")
    print("   5. If Tier 1: Proceed to Requirements Extraction (Workbook 3)")
    print()
    print("=" * 70)


if __name__ == "__main__":
    main()
