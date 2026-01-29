#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.12.4 - Monitoring & Response Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Annex A Control A.8.12: Data Leakage Prevention
Assessment Domain 4 of 4: Monitoring & Response

--------------------------------------------------------------------------------
PURPOSE
--------------------------------------------------------------------------------

Generates comprehensive monitoring and response assessment workbook for systematic
evaluation of DLP detection, alerting, and incident response capabilities against
ISO 27001:2022 Control A.8.12 requirements.

This workbook provides audit-ready evidence collection framework covering:
• Logging configuration for all DLP channels
• Alert rules inventory and alert volume metrics
• SIEM integration and log correlation
• False positive management and tuning procedures
• Incident response workflow and SLA compliance
• SOC integration and escalation procedures
• Executive and operational dashboards
• Gap analysis and remediation planning
• Evidence register for audit traceability

--------------------------------------------------------------------------------
GENERATED WORKBOOK STRUCTURE
--------------------------------------------------------------------------------

Output File: ISMS-IMP-A.8.12.4_Monitoring_Response_YYYYMMDD.xlsx

Sheets (12 total):
1. Instructions_Legend - Assessment guidance and metadata
2. Logging_Configuration - DLP logging requirements per channel
3. Alert_Rules_Inventory - All configured DLP alert rules
4. Alert_Volume_Metrics - Daily/weekly/monthly alert statistics
5. SIEM_Integration - Log forwarding, correlation rules, use cases
6. False_Positive_Management - FP rate tracking, tuning procedures
7. Incident_Response_Workflow - Triage, containment, investigation
8. SOC_Integration - Alert workflow, escalation, SLAs
9. Dashboards_Reporting - Executive, operational, compliance dashboards
10. Gap_Analysis - Monitoring and response gaps (40 rows)
11. Evidence_Register - Audit evidence tracking (100 rows)
12. Summary_Dashboard - Compliance metrics and KPIs

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System:
    Python 3.8 or higher
    Ubuntu/Debian Linux (recommended) or macOS

Dependencies:
    openpyxl - Excel file generation library
    
Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic:
    python3 generate_a812_4_monitoring_response.py

Output Location:
    Current working directory
    
Output Filename:
    ISMS-IMP-A.8.12.4_Monitoring_Response_YYYYMMDD.xlsx
    (Where YYYYMMDD = current date)

Post-Generation:
    1. Open workbook in Microsoft Excel or LibreOffice Calc
    2. Complete all yellow input fields (organization-specific data)
    3. Review pre-populated examples (gray rows) for guidance
    4. Document all DLP logging configurations
    5. Inventory alert rules and analyze alert volumes
    6. Assess SIEM integration and correlation capabilities
    7. Measure false positive rates and tuning effectiveness
    8. Verify incident response procedures and SLA compliance
    9. Collect and document evidence (Evidence_Register sheet)
    10. Complete gap analysis for monitoring deficiencies
    11. Obtain management approval (Summary_Dashboard sheet)

--------------------------------------------------------------------------------
FRAMEWORK INTEGRATION
--------------------------------------------------------------------------------

Position in A.8.12 Framework:
    Assessment Domain 4 of 4 in comprehensive DLP evaluation framework
    Focus: Detection, monitoring, alerting, and incident response
    
Related Documents:
    Policy:         ISMS-POL-A.8.12-S2.3 (Monitoring & Detection Requirements)
    Policy:         ISMS-POL-A.8.12-S2.4 (Incident Response & Remediation)
    Policy:         ISMS-POL-A.8.12-S5.C (DLP Incident Response Procedures)
    Implementation: ISMS-IMP-A.8.12.4 (Monitoring & Response Implementation Guide)
    Dashboard:      ISMS-IMP-A.8.12.5 (Compliance Dashboard)

Integration Workflow:
    1. Generate assessment workbooks:
       python3 generate_a812_1_dlp_infrastructure.py
       python3 generate_a812_2_data_classification.py
       python3 generate_a812_3_channel_coverage.py
       python3 generate_a812_4_monitoring_response.py      ← YOU ARE HERE
    
    2. Complete assessments (manual - security team, SOC analysts)
    
    3. Normalize filenames for dashboard linking:
       python3 normalize_assessment_files_a812.py
    
    4. Generate executive dashboard:
       python3 generate_a812_5_compliance_dashboard.py
    
    5. Consolidate assessment data:
       python3 consolidate_a812_dashboard.py [dashboard_file]
    
    6. Present to CISO/auditors (evidence-based compliance reporting)

Data Flow:
    THIS SCRIPT → Monitoring Assessment → Normalize → Dashboard → Audit Evidence

Critical Prerequisites:
    • Domain 1 (Infrastructure) - need DLP technology capabilities
    • Domain 3 (Channel Coverage) - need alert rules per channel
    • SIEM/SOC team involvement for integration assessment

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.12
Assessment Domain:    4 of 4 (Monitoring & Response)
Framework Version:    1.0
Script Version:       1.0
Date:                 25.01.2025
Author:               [Organization] ISMS Implementation Team
License:              [Organization License/Terms]

Related Standards:
    - ISO/IEC 27002:2022 (Information Security Controls)
    - ISO/IEC 27035:2023 (Information Security Incident Management)
    - Swiss FADP (Federal Act on Data Protection - Breach Notification)
    - EU GDPR (General Data Protection Regulation - Articles 33, 34)
    - NIST SP 800-53 (Security and Privacy Controls - IR, SI families)
    - NIST SP 800-61 (Computer Security Incident Handling Guide)
    - CIS Controls v8 (Center for Internet Security - Controls 8, 17)

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

Alert Severity Classification (from ISMS-POL-A.8.12-S2.3):
    Critical (Response SLA: 15 minutes):
        • Active data exfiltration detected
        • Credentials (passwords, API keys) detected in transit
        • Large-volume sensitive data transfers
    
    High (Response SLA: 1 hour):
        • PII/IP transferred externally without encryption
        • Sensitive data sent to unauthorized cloud services
        • Policy violations by privileged users
    
    Medium (Response SLA: 4 hours):
        • Confidential data transfers requiring review
        • Repeated policy violations by single user
        • Suspicious patterns requiring investigation
    
    Low (Response SLA: 24 hours):
        • Policy violations requiring tuning
        • User education opportunities
        • Process improvement notifications
    
    Informational (No SLA):
        • Monitoring-only events
        • Audit trail entries
        • Compliance reporting data

Swiss FADP Breach Notification Requirements:
    Data breach notification obligations under Swiss Federal Act on Data Protection:
    • Notification to Federal Data Protection and Information Commissioner (FDPIC)
    • Notification required when breach likely to result in high risk to data subjects
    • Notification must be "without delay" (typically within 72 hours)
    • DLP monitoring must detect breaches in time to meet notification deadlines
    
    This assessment verifies:
    • DLP alerts configured to detect breach scenarios
    • Incident response procedures align with FADP timelines
    • Evidence collection supports breach notification reporting

Assessment Scope - Monitoring Components:
    1. Logging Configuration:
       • Log all DLP events (alerts, policy matches, data transfers)
       • Retention period: minimum 90 days (recommend 1 year)
       • Log format: structured (JSON/CEF) for SIEM ingestion
       • Protected logs: tamper-proof, encrypted storage
    
    2. Alert Rules:
       • Content inspection rules (keywords, regex patterns)
       • Data classification label rules
       • File type and size threshold rules
       • User/group exception rules
       • Contextual rules (time, location, device)
    
    3. SIEM Integration:
       • Real-time log forwarding (syslog, API)
       • Correlation rules (DLP + network + endpoint)
       • Use cases (insider threat, data exfiltration)
       • Alert enrichment (user context, asset context)
    
    4. False Positive Management:
       • FP rate measurement (target: <5%)
       • Root cause analysis procedures
       • Policy tuning workflows
       • Exception management process
    
    5. Incident Response:
       • Triage procedures (severity classification)
       • Containment actions (block user, quarantine file)
       • Investigation steps (forensics, user interview)
       • Remediation and lessons learned

Key Performance Indicators (KPIs):
    • Alert Volume: Daily/weekly/monthly trends
    • False Positive Rate: Target <5%
    • Mean Time to Detect (MTTD): Target <1 hour for Critical
    • Mean Time to Respond (MTTR): Per severity SLA
    • Incident Closure Rate: % within SLA
    • Tuning Effectiveness: Reduction in FP rate over time

Data Classification:
    Generated workbooks contain sensitive organizational security information.
    Handle according to [Organization]'s data classification policy.
    Recommended classification: [Organization] Internal/Confidential

Audit Considerations:
    This workbook generates ISO 27001:2022 audit evidence per Control A.8.12.
    Ensure all fields completed accurately and evidence properly documented.
    Retain completed workbooks for audit cycle (typically 3 years).
    Auditors will verify: logging completeness, SIEM integration, incident response.

Review Cycle:
    Monthly: Review alert volumes and false positive rates
    Quarterly: Update monitoring configuration and tuning status
    Annually: Complete full monitoring and response reassessment
    Ad-hoc: After security incidents or infrastructure changes

================================================================================
END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •



# ============================================================================
# SECTION 1: CONSTANTS & CONFIGURATION
# ============================================================================

# Document Information
WORKBOOK_VERSION = "1.0"
CONTROL_ID = "A.8.12"
WORKBOOK_ID = "ISMS-IMP-A.8.12.4"
RELATED_POLICY = "ISMS-POL-A.8.12-S2.3, ISMS-POL-A.8.12-S2.4"
ASSESSMENT_AREA = "Monitoring & Response Assessment"

# Color Scheme (CONSISTENT across all A.8.12 workbooks)
COLOR_HEADER = "003366"          # Dark blue
COLOR_SUBHEADER = "4472C4"       # Medium blue
COLOR_COLUMN_HEADER = "D9D9D9"   # Light gray
COLOR_INPUT = "FFFFCC"           # Light yellow (user input)
COLOR_INFO = "E7E6E6"            # Light gray (info/example rows)
COLOR_COMPLETE = "C6EFCE"        # Light green (\u2705 Yes)
COLOR_PARTIAL = "FFEB9C"         # Light yellow (\u26A0\uFE0F Partial)
COLOR_MISSING = "FFC7CE"         # Light red (\u274C No)
COLOR_PLANNED = "B4C7E7"         # Light blue (\u1F4CB Planned)

# Standard column widths
WIDTH_NARROW = 12
WIDTH_MEDIUM = 20
WIDTH_WIDE = 25
WIDTH_EXTRA_WIDE = 30
WIDTH_DESCRIPTION = 40


# ============================================================================
# SECTION 2: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create all sheets in order
    sheets = [
        "Instructions_Legend",
        "Logging_Configuration",
        "Alert_Rules_Inventory",
        "Alert_Volume_Metrics",
        "SIEM_Integration",
        "False_Positive_Management",
        "Incident_Response_Workflow",
        "SOC_Integration",
        "Dashboards_Reporting",
        "Gap_Analysis",
        "Evidence_Register",
        "Summary_Dashboard",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    
    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style TEMPLATES (dictionaries), not reusable objects.
    
    CRITICAL: Do NOT create shared Font/Fill/Border objects.
    Each cell gets its OWN style instance to avoid openpyxl issues.
    """
    return {
        "header": {
            "font": Font(name="Arial", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "subheader": {
            "font": Font(name="Arial", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "column_header": {
            "font": Font(name="Arial", size=11, bold=True),
            "fill": PatternFill(start_color=COLOR_COLUMN_HEADER, end_color=COLOR_COLUMN_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "input_cell": {
            "font": Font(name="Arial", size=10),
            "fill": PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "info_cell": {
            "font": Font(name="Arial", size=10, italic=True),
            "fill": PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "data_cell": {
            "font": Font(name="Arial", size=10),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
    }


def apply_style(cell, style_dict):
    """
    Apply style template to cell by creating NEW instances.
    
    CRITICAL: Each cell gets its own Font/Fill/Border/Alignment objects.
    This prevents openpyxl's "Cannot use already used style object" error.
    """
    if "font" in style_dict:
        f = style_dict["font"]
        cell.font = Font(
            name=f.name,
            size=f.size,
            bold=f.bold,
            italic=f.italic,
            color=f.color
        )
    
    if "fill" in style_dict:
        f = style_dict["fill"]
        cell.fill = PatternFill(
            start_color=f.start_color.rgb if f.start_color else None,
            end_color=f.end_color.rgb if f.end_color else None,
            fill_type=f.fill_type
        )
    
    if "alignment" in style_dict:
        a = style_dict["alignment"]
        cell.alignment = Alignment(
            horizontal=a.horizontal,
            vertical=a.vertical,
            wrap_text=a.wrap_text
        )
    
    if "border" in style_dict:
        b = style_dict["border"]
        cell.border = Border(
            left=Side(style=b.left.style) if b.left else None,
            right=Side(style=b.right.style) if b.right else None,
            top=Side(style=b.top.style) if b.top else None,
            bottom=Side(style=b.bottom.style) if b.bottom else None
        )


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def setup_data_validations(wb):
    """Setup all data validation dropdowns used in the workbook."""
    
    # Standard response values (Yes/No/Partial/Planned/N/A)
    val_response = DataValidation(
        type="list",
        formula1='"Yes,No,Partial,Planned,N/A"',
        allow_blank=False
    )
    
    # Alert severity levels
    val_severity = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low,Informational"',
        allow_blank=False
    )
    
    # Channel types
    val_channel = DataValidation(
        type="list",
        formula1='"Email,Web,Endpoint,Network,Application,Mobile"',
        allow_blank=False
    )
    
    # Detection methods
    val_detection = DataValidation(
        type="list",
        formula1='"Pattern,Keyword,Fingerprint,Contextual,ML"',
        allow_blank=False
    )
    
    # Policy actions
    val_action = DataValidation(
        type="list",
        formula1='"Allow,Alert,Block,Quarantine,Encrypt"',
        allow_blank=False
    )
    
    # Alert destinations
    val_destination = DataValidation(
        type="list",
        formula1='"SIEM,Email,SMS,Ticketing,Dashboard"',
        allow_blank=False
    )
    
    # Response SLAs
    val_sla = DataValidation(
        type="list",
        formula1='"15min,1hr,4hr,24hr,No SLA"',
        allow_blank=False
    )
    
    # Rule status
    val_rule_status = DataValidation(
        type="list",
        formula1='"Active,Disabled,Testing,Deprecated"',
        allow_blank=False
    )
    
    # Evidence types
    val_evidence = DataValidation(
        type="list",
        formula1='"Config,Screenshot,Log,Report,Dashboard,Other"',
        allow_blank=False
    )
    
    # Verification status
    val_verification = DataValidation(
        type="list",
        formula1='"Verified,Pending,Rejected"',
        allow_blank=False
    )
    
    # Risk levels
    val_risk = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    
    # Gap status
    val_gap_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Complete,Blocked"',
        allow_blank=False
    )
    
    # Affected areas for gaps
    val_affected_area = DataValidation(
        type="list",
        formula1='"Logging,Alerting,SIEM,FP Management,Incident Response,SOC,Dashboards"',
        allow_blank=False
    )
    
    return {
        "response": val_response,
        "severity": val_severity,
        "channel": val_channel,
        "detection": val_detection,
        "action": val_action,
        "destination": val_destination,
        "sla": val_sla,
        "rule_status": val_rule_status,
        "evidence": val_evidence,
        "verification": val_verification,
        "risk": val_risk,
        "gap_status": val_gap_status,
        "affected_area": val_affected_area,
    }


# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 1
# ============================================================================

def create_instructions_sheet(wb, styles):
    """Create Instructions_Legend sheet."""
    ws = wb["Instructions_Legend"]
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = f"{WORKBOOK_ID} - {ASSESSMENT_AREA}"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Instructions & Legend"
    apply_style(ws['A2'], styles["subheader"])
    
    # Document metadata
    row = 4
    metadata = [
        ("Workbook ID:", WORKBOOK_ID),
        ("Assessment Area:", ASSESSMENT_AREA),
        ("Related Policies:", RELATED_POLICY),
        ("Version:", WORKBOOK_VERSION),
        ("ISO Control:", CONTROL_ID),
    ]
    
    for label, value in metadata:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    # Organization metadata section
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "ORGANIZATION METADATA (Complete the yellow fields)"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    org_fields = [
        "Assessment Date:",
        "Completed By (Name):",
        "Completed By (Role):",
        "Organization Name:",
        "Review Cycle:",
    ]
    
    for field in org_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(bold=True)
        apply_style(ws[f'B{row}'], styles["input_cell"])
        if field == "Review Cycle:":
            ws[f'B{row}'] = "Quarterly"
        row += 1
    
    # How to use section
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "HOW TO USE THIS WORKBOOK"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    instructions = [
        "1. Complete Logging_Configuration first (foundational)",
        "2. Document all Alert_Rules_Inventory (what alerts exist)",
        "3. Track Alert_Volume_Metrics (daily/weekly/monthly statistics)",
        "4. Verify SIEM_Integration (log forwarding, correlation rules)",
        "5. Measure False_Positive_Management (FP rate, tuning process)",
        "6. Document Incident_Response_Workflow (triage, SLA compliance)",
        "7. Assess SOC_Integration (alert workflow, escalation)",
        "8. Verify Dashboards_Reporting (executive, operational dashboards)",
        "9. Identify gaps and create remediation plans in Gap_Analysis",
        "10. Collect evidence and document in Evidence_Register",
        "11. Review Summary_Dashboard for overall compliance score",
        "",
        "Evidence ID Format: A812-4-[CATEGORY]-[###]",
        "Examples: A812-4-LOG-001, A812-4-ALT-001, A812-4-SIEM-001",
    ]
    
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Legend section
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "LEGEND - RESPONSE VALUES"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    legend_items = [
        ("\u2705 Yes", "Fully implemented, documented, measured with evidence"),
        ("\u274C No", "Not implemented or significant gaps exist"),
        ("\u26A0\uFE0F Partial", "Partially implemented, some gaps remain"),
        ("\u1F4CB Planned", "Scheduled for implementation (provide target date)"),
        ("⚪ N/A", "Not applicable (provide justification)"),
        ("❗ FORBIDDEN", "\"Maybe\" is NOT allowed (Feynman principle: either you know or you don't)"),
    ]
    
    for symbol, description in legend_items:
        ws[f'A{row}'] = symbol
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = description
        ws.merge_cells(f'B{row}:E{row}')
        row += 1
    
    # Alert severity classification
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "ALERT SEVERITY CLASSIFICATION (from ISMS-POL-A.8.12-S2.3)"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    severity_data = [
        ("Critical", "Active data exfiltration, credentials detected", "15 minutes"),
        ("High", "PII/IP transferred externally without encryption", "1 hour"),
        ("Medium", "Confidential data transfer requiring review", "4 hours"),
        ("Low", "Policy violation, tuning needed", "24 hours"),
        ("Informational", "Monitoring-only events", "No SLA"),
    ]
    
    # Header row
    ws[f'A{row}'] = "Severity"
    ws[f'B{row}'] = "Definition"
    ws[f'C{row}'] = "Response SLA"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f'{col}{row}'], styles["column_header"])
    row += 1
    
    for severity, definition, sla in severity_data:
        ws[f'A{row}'] = severity
        ws[f'B{row}'] = definition
        ws[f'C{row}'] = sla
        ws.merge_cells(f'B{row}:D{row}')
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A3'


def create_logging_configuration_sheet(wb, styles, validations):
    """Create Logging_Configuration sheet."""
    ws = wb["Logging_Configuration"]
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "DLP Logging Configuration Assessment"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 25
    
    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Assess DLP logging requirements per channel (25 criteria)"
    apply_style(ws['A2'], styles["subheader"])
    
    # Column headers
    headers = ["#", "Requirement", "Compliance", "Evidence ID", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Assessment criteria
    requirements = [
        "Are all DLP channels configured to log events?",
        "Email DLP: Logs sent/received/blocked emails with metadata?",
        "Web DLP: Logs uploads/downloads/blocked web traffic?",
        "Endpoint DLP: Logs USB/print/screen capture/clipboard events?",
        "Network DLP: Logs file transfers (FTP/SMB/SCP)?",
        "Application DLP: Logs database exports/API calls?",
        "Mobile DLP: Logs mobile app usage and data transfers?",
        "Do logs include user identity (username, email, employee ID)?",
        "Do logs include timestamp (ISO 8601 format)?",
        "Do logs include source (IP, hostname, device ID)?",
        "Do logs include destination (recipient, URL, file path)?",
        "Do logs include data classification level (if available)?",
        "Do logs include policy violated (rule ID, rule name)?",
        "Do logs include action taken (allow/alert/block/quarantine)?",
        "Do logs include file metadata (filename, size, hash)?",
        "Are logs retained for minimum 12 months (FADP requirement)?",
        "Are logs protected from unauthorized modification?",
        "Are logs encrypted in transit (TLS to SIEM)?",
        "Are logs encrypted at rest (SIEM storage)?",
        "Is log integrity verified (checksums, digital signatures)?",
        "Are logs reviewed for completeness (missing log gaps)?",
        "Are logs backed up (separate from production logs)?",
        "Is log access restricted (RBAC, audit trail of log access)?",
        "Are logs parsed correctly by SIEM (no parsing errors)?",
        "Are logs correlated with other security events (IAM, EDR, Network)?",
    ]
    
    row = 4
    # Add 3 example rows first (gray)
    example_data = [
        (1, requirements[0], "Yes", "A812-4-LOG-001", "All channels logging to Splunk"),
        (2, requirements[1], "Partial", "A812-4-LOG-002", "Missing encrypted email logs"),
        (3, requirements[2], "Yes", "A812-4-LOG-003", "Web proxy logs complete"),
    ]
    
    for num, req, status, evidence, notes in example_data:
        ws[f'A{row}'] = num
        ws[f'B{row}'] = req
        ws[f'C{row}'] = status
        ws[f'D{row}'] = evidence
        ws[f'E{row}'] = notes
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["info_cell"])
        
        row += 1
    
    # Add remaining blank rows
    for idx, req in enumerate(requirements, start=1):
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = req
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["data_cell"])
        
        row += 1
    
    # Add data validation to Compliance column (starting after examples)
    val_response = validations["response"]
    ws.add_data_validation(val_response)
    for r in range(7, 29):  # Rows 7-28 (after 3 examples)
        val_response.add(ws[f'C{r}'])
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_NARROW
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_alert_rules_inventory_sheet(wb, styles, validations):
    """Create Alert_Rules_Inventory sheet."""
    ws = wb["Alert_Rules_Inventory"]
    
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "DLP Alert Rules Inventory"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 25
    
    # Subheader
    ws.merge_cells('A2:M2')
    ws['A2'] = "Complete inventory of all configured DLP alert rules (30 rows)"
    apply_style(ws['A2'], styles["subheader"])
    
    # Column headers
    headers = [
        "Alert Rule ID", "Rule Name", "Channel", "Severity", "Detection Method",
        "Policy Action", "Alert Destination", "Response SLA", "Last Tuned Date",
        "FP Rate %", "Alert Volume (30d)", "Rule Status", "Evidence ID"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Pre-populated examples (5 rows, gray)
    example_rules = [
        ("ALT-001", "Credentials Detected - Email", "Email", "Critical", "Pattern", 
         "Block", "SIEM", "15min", "2025-01-01", 2.1, 45, "Active", "A812-4-ALT-001"),
        ("ALT-002", "PII to External Domain", "Email", "High", "Fingerprint",
         "Alert", "SIEM", "1hr", "2024-12-15", 8.5, 230, "Active", "A812-4-ALT-002"),
        ("ALT-003", "Source Code to GitHub", "Web", "High", "Keyword",
         "Block", "SIEM", "1hr", "2025-01-03", 1.2, 12, "Active", "A812-4-ALT-003"),
        ("ALT-004", "Mass USB Copy (>1GB)", "Endpoint", "Critical", "Contextual",
         "Block", "SIEM", "15min", "2024-11-20", 0.5, 3, "Active", "A812-4-ALT-004"),
        ("ALT-005", "Bulk Database Export", "Application", "High", "ML",
         "Alert", "SIEM", "1hr", "2024-12-28", 4.3, 67, "Active", "A812-4-ALT-005"),
    ]
    
    row = 4
    for rule_data in example_rules:
        for col_idx, value in enumerate(rule_data, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # Add blank rows (25 more = 30 total)
    for _ in range(25):
        for col_idx in range(1, 14):
            cell = ws.cell(row=row, column=col_idx)
            apply_style(cell, styles["data_cell"])
        row += 1
    
    # Add data validations
    val_channel = validations["channel"]
    val_severity = validations["severity"]
    val_detection = validations["detection"]
    val_action = validations["action"]
    val_destination = validations["destination"]
    val_sla = validations["sla"]
    val_rule_status = validations["rule_status"]
    
    ws.add_data_validation(val_channel)
    ws.add_data_validation(val_severity)
    ws.add_data_validation(val_detection)
    ws.add_data_validation(val_action)
    ws.add_data_validation(val_destination)
    ws.add_data_validation(val_sla)
    ws.add_data_validation(val_rule_status)
    
    for r in range(9, 34):  # Rows 9-33 (after 5 examples)
        val_channel.add(ws[f'C{r}'])
        val_severity.add(ws[f'D{r}'])
        val_detection.add(ws[f'E{r}'])
        val_action.add(ws[f'F{r}'])
        val_destination.add(ws[f'G{r}'])
        val_sla.add(ws[f'H{r}'])
        val_rule_status.add(ws[f'L{r}'])
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    ws.column_dimensions['F'].width = WIDTH_MEDIUM
    ws.column_dimensions['G'].width = WIDTH_MEDIUM
    ws.column_dimensions['H'].width = WIDTH_NARROW
    ws.column_dimensions['I'].width = WIDTH_MEDIUM
    ws.column_dimensions['J'].width = WIDTH_NARROW
    ws.column_dimensions['K'].width = WIDTH_MEDIUM
    ws.column_dimensions['L'].width = WIDTH_MEDIUM
    ws.column_dimensions['M'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A4'

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 2
# ============================================================================

def create_alert_volume_metrics_sheet(wb, styles):
    """Create Alert_Volume_Metrics sheet."""
    ws = wb["Alert_Volume_Metrics"]
    
    # Header
    ws.merge_cells('A1:G1')
    ws['A1'] = "DLP Alert Volume Metrics"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 25
    
    # Subheader
    ws.merge_cells('A2:G2')
    ws['A2'] = "Track alert volume by severity and channel (Last 30 Days)"
    apply_style(ws['A2'], styles["subheader"])
    
    row = 4
    
    # Section 1: Alert Volume by Severity
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "ALERT VOLUME BY SEVERITY (Last 30 Days)"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    severity_headers = ["Severity", "Total Alerts", "Avg per Day", "Blocked", "Alerted", "False Positives", "FP Rate %"]
    for col_idx, header in enumerate(severity_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    severities = ["Critical", "High", "Medium", "Low", "Informational", "TOTAL"]
    for sev in severities:
        ws[f'A{row}'] = sev
        if sev == "TOTAL":
            ws[f'A{row}'].font = Font(bold=True)
            # Formulas for TOTAL row
            ws[f'B{row}'] = f'=SUM(B{row-5}:B{row-1})'
            ws[f'C{row}'] = f'=B{row}/30'
            ws[f'D{row}'] = f'=SUM(D{row-5}:D{row-1})'
            ws[f'E{row}'] = f'=SUM(E{row-5}:E{row-1})'
            ws[f'F{row}'] = f'=SUM(F{row-5}:F{row-1})'
            ws[f'G{row}'] = f'=IF(B{row}>0,F{row}/B{row}*100,0)'
        else:
            # Input cells for data entry
            for col in ['B', 'C', 'D', 'E', 'F']:
                apply_style(ws[f'{col}{row}'], styles["input_cell"])
            # FP Rate formula
            ws[f'G{row}'] = f'=IF(B{row}>0,F{row}/B{row}*100,0)'
        row += 1
    
    row += 1
    
    # Section 2: Alert Volume by Channel
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "ALERT VOLUME BY CHANNEL (Last 30 Days)"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    channel_headers = ["Channel", "Total Alerts", "Blocked", "Alerted", "FP Rate %", "Top Rule Name"]
    for col_idx, header in enumerate(channel_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    channels = ["Email", "Web", "Endpoint", "Network", "Application", "Mobile"]
    for channel in channels:
        ws[f'A{row}'] = channel
        for col in ['B', 'C', 'D', 'E', 'F']:
            apply_style(ws[f'{col}{row}'], styles["input_cell"])
        row += 1
    
    row += 1
    
    # Section 3: Alert Trend Analysis (4 weeks)
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "ALERT TREND ANALYSIS (Weekly)"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    trend_headers = ["Week", "Critical", "High", "Medium", "Low", "Total", "FP Rate %", "Notes"]
    for col_idx, header in enumerate(trend_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    weeks = ["Week 1 (Dec 2-8)", "Week 2 (Dec 9-15)", "Week 3 (Dec 16-22)", "Week 4 (Dec 23-29)"]
    for week in weeks:
        ws[f'A{row}'] = week
        # Input cells
        for col in ['B', 'C', 'D', 'E', 'G', 'H']:
            apply_style(ws[f'{col}{row}'], styles["input_cell"])
        # Total formula
        ws[f'F{row}'] = f'=SUM(B{row}:E{row})'
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    ws.column_dimensions['F'].width = WIDTH_MEDIUM
    ws.column_dimensions['G'].width = WIDTH_MEDIUM
    ws.column_dimensions['H'].width = WIDTH_EXTRA_WIDE
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_siem_integration_sheet(wb, styles, validations):
    """Create SIEM_Integration sheet."""
    ws = wb["SIEM_Integration"]
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "SIEM Integration Assessment"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 25
    
    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Assess SIEM integration, log forwarding, and correlation rules (25 criteria)"
    apply_style(ws['A2'], styles["subheader"])
    
    # Column headers
    headers = ["#", "Requirement", "Compliance", "Evidence ID", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Assessment criteria
    requirements = [
        "Are DLP logs forwarded to organizational SIEM?",
        "Is log forwarding encrypted (TLS/SSL)?",
        "Is log forwarding reliable (no dropped logs >1%)?",
        "Are DLP logs parsed correctly in SIEM (no parsing errors)?",
        "Are DLP logs normalized (common schema: CEF, Syslog, JSON)?",
        "Are DLP events indexed for search and analysis?",
        "Is log retention in SIEM ≥12 months?",
        "Are SIEM correlation rules configured for DLP events?",
        "Correlation: DLP + Malware detection = compromised endpoint?",
        "Correlation: DLP + Failed logins = credential theft?",
        "Correlation: DLP + VPN from foreign country = account compromise?",
        "Correlation: DLP + HR termination record = insider threat?",
        "Are SIEM use cases documented (insider threat, compromised creds)?",
        "Are SIEM dashboards configured for DLP (real-time alert queue)?",
        "Are SIEM alerts integrated with ticketing system (ServiceNow, Jira)?",
        "Are DLP Critical alerts automatically escalated to SOC?",
        "Are DLP High alerts automatically assigned to analysts?",
        "Is DLP log ingestion monitored (alerts for log failures)?",
        "Is SIEM performance adequate (query response time <10 sec)?",
        "Are SIEM search queries documented (common investigations)?",
        "Is SIEM access restricted (RBAC, SOC analysts only)?",
        "Are SIEM searches audited (who searched for what)?",
        "Is SIEM backup configured (disaster recovery)?",
        "Are SIEM correlation rules tested (validation, false positive rate)?",
        "Are SIEM correlation rules tuned quarterly?",
    ]
    
    row = 4
    # Add 3 example rows (gray)
    example_data = [
        (1, requirements[0], "Yes", "A812-4-SIEM-001", "Forwarding to Splunk via TLS"),
        (2, requirements[7], "Partial", "A812-4-SIEM-002", "Only 3 of 6 use cases deployed"),
        (3, requirements[13], "Yes", "A812-4-SIEM-003", "SOC dashboard configured"),
    ]
    
    for num, req, status, evidence, notes in example_data:
        ws[f'A{row}'] = num
        ws[f'B{row}'] = req
        ws[f'C{row}'] = status
        ws[f'D{row}'] = evidence
        ws[f'E{row}'] = notes
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["info_cell"])
        
        row += 1
    
    # Add remaining blank rows
    for idx, req in enumerate(requirements, start=1):
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = req
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["data_cell"])
        
        row += 1
    
    # Add data validation
    val_response = validations["response"]
    ws.add_data_validation(val_response)
    for r in range(7, 29):
        val_response.add(ws[f'C{r}'])
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_NARROW
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_false_positive_management_sheet(wb, styles, validations):
    """Create False_Positive_Management sheet."""
    ws = wb["False_Positive_Management"]
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "False Positive Management Assessment"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 25
    
    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Assess false positive tracking, tuning, and management (20 criteria + FP rate table)"
    apply_style(ws['A2'], styles["subheader"])
    
    # Column headers
    headers = ["#", "Requirement", "Compliance", "Evidence ID", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Assessment criteria
    requirements = [
        "Is false positive rate measured per alert rule?",
        "Is overall FP rate <10% (target from policy)?",
        "Are false positives tracked in ticketing system?",
        "Is FP rate trended over time (monthly)?",
        "Are users able to report false positives (self-service portal)?",
        "Are false positive reports triaged within 24 hours?",
        "Are false positives investigated (root cause analysis)?",
        "Are DLP rules tuned based on FP analysis?",
        "Are legitimate business flows whitelisted (permanent exceptions)?",
        "Are whitelists documented (justification, approver, review date)?",
        "Are whitelists reviewed annually (remove stale exceptions)?",
        "Are high-FP rules disabled or re-tuned?",
        "Are users educated when FP is user error (not DLP issue)?",
        "Is FP tuning tracked (before/after FP rate improvement)?",
        "Are tuning changes tested before production deployment?",
        "Are tuning changes documented (change request, approval)?",
        "Are tuning changes rolled back if new issues arise?",
        "Is FP rate included in SOC performance metrics?",
        "Are FP trends reported to management (monthly)?",
        "Are FP reduction targets set and tracked?",
    ]
    
    row = 4
    # Add 3 example rows
    example_data = [
        (1, requirements[0], "Yes", "A812-4-FP-001", "FP rate tracked per rule in dashboard"),
        (2, requirements[1], "Partial", "A812-4-FP-002", "Overall 8.7%, close to 10% target"),
        (3, requirements[7], "Yes", "A812-4-FP-003", "Monthly tuning sessions scheduled"),
    ]
    
    for num, req, status, evidence, notes in example_data:
        ws[f'A{row}'] = num
        ws[f'B{row}'] = req
        ws[f'C{row}'] = status
        ws[f'D{row}'] = evidence
        ws[f'E{row}'] = notes
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["info_cell"])
        
        row += 1
    
    # Add remaining blank rows
    for idx, req in enumerate(requirements, start=1):
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = req
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["data_cell"])
        
        row += 1
    
    # Add data validation
    val_response = validations["response"]
    ws.add_data_validation(val_response)
    for r in range(7, 27):
        val_response.add(ws[f'C{r}'])
    
    row += 1
    
    # FP Rate Calculation Table
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "FALSE POSITIVE RATE BY ALERT RULE (Top 10 Rules)"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    fp_headers = ["Alert Rule ID", "Rule Name", "Total Alerts (30d)", "False Positives", "FP Rate %", "Target FP %", "Status"]
    for col_idx, header in enumerate(fp_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    # Example rows
    fp_examples = [
        ("ALT-001", "Credentials - Email", 45, 1, "=D{}/C{}", "5%", "\u2705 On Target"),
        ("ALT-002", "PII to External", 230, 20, "=D{}/C{}", "10%", "\u2705 On Target"),
        ("ALT-003", "Source Code GitHub", 12, 0, "=D{}/C{}", "5%", "\u2705 On Target"),
    ]
    
    for rule_id, rule_name, total, fp, formula, target, status in fp_examples:
        ws[f'A{row}'] = rule_id
        ws[f'B{row}'] = rule_name
        ws[f'C{row}'] = total
        ws[f'D{row}'] = fp
        ws[f'E{row}'] = formula.format(row, row)
        ws[f'F{row}'] = target
        ws[f'G{row}'] = status
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            apply_style(ws[f'{col}{row}'], styles["info_cell"])
        
        row += 1
    
    # Blank rows for additional entries
    for _ in range(7):
        for col in ['A', 'B', 'C', 'D', 'F', 'G']:
            apply_style(ws.cell(row=row, column=ord(col)-ord('A')+1), styles["input_cell"])
        ws[f'E{row}'] = f'=IF(C{row}>0,D{row}/C{row}*100,0)'
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_incident_response_workflow_sheet(wb, styles, validations):
    """Create Incident_Response_Workflow sheet."""
    ws = wb["Incident_Response_Workflow"]
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "Incident Response Workflow Assessment"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 25
    
    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Assess incident response procedures and SLA compliance (25 criteria + SLA tracking)"
    apply_style(ws['A2'], styles["subheader"])
    
    # Column headers
    headers = ["#", "Requirement", "Compliance", "Evidence ID", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Assessment criteria
    requirements = [
        "Are DLP alerts integrated with incident response workflow?",
        "Is incident response procedure documented (playbook)?",
        "Critical alerts: Are alerts triaged within 15 minutes?",
        "High alerts: Are alerts triaged within 1 hour?",
        "Medium alerts: Are alerts triaged within 4 hours?",
        "Low alerts: Are alerts triaged within 24 hours?",
        "Is alert triage documented (who, when, decision)?",
        "Are false positives closed with justification?",
        "Are true positives escalated to incident response team?",
        "Are users contacted for business justification (High alerts)?",
        "Are user managers contacted for approval (policy exceptions)?",
        "Are incidents assigned severity based on data classification?",
        "Are incidents assigned to appropriate team (SOC, IR, Legal)?",
        "Is containment performed (block user, disable account, isolate endpoint)?",
        "Is investigation performed (correlate with other events)?",
        "Is evidence collected (logs, screenshots, file samples)?",
        "Is evidence preserved (chain of custody)?",
        "Are incidents reported to management (Critical/High)?",
        "Are incidents reported to DPO (personal data breaches)?",
        "Are incidents reported to CISO (all Critical incidents)?",
        "Are incidents reported to Legal (potential litigation)?",
        "Is remediation tracked (close incident when resolved)?",
        "Is Mean Time to Detect (MTTD) measured?",
        "Is Mean Time to Respond (MTTR) measured?",
        "Are lessons learned documented (post-incident review)?",
    ]
    
    row = 4
    # Add 3 example rows
    example_data = [
        (1, requirements[0], "Yes", "A812-4-IR-001", "ServiceNow integration complete"),
        (2, requirements[2], "Partial", "A812-4-IR-002", "SLA 87.5%, need improvement"),
        (3, requirements[22], "Yes", "A812-4-IR-003", "MTTD tracked: avg 8 minutes"),
    ]
    
    for num, req, status, evidence, notes in example_data:
        ws[f'A{row}'] = num
        ws[f'B{row}'] = req
        ws[f'C{row}'] = status
        ws[f'D{row}'] = evidence
        ws[f'E{row}'] = notes
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["info_cell"])
        
        row += 1
    
    # Add remaining blank rows
    for idx, req in enumerate(requirements, start=1):
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = req
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["data_cell"])
        
        row += 1
    
    # Add data validation
    val_response = validations["response"]
    ws.add_data_validation(val_response)
    for r in range(7, 32):
        val_response.add(ws[f'C{r}'])
    
    row += 1
    
    # SLA Compliance Tracking Table
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "SLA COMPLIANCE TRACKING (Last 30 Days)"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    sla_headers = ["Severity", "SLA Target", "Incidents (30d)", "Within SLA", "SLA Compliance %"]
    for col_idx, header in enumerate(sla_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    row += 1
    
    # Example rows
    sla_examples = [
        ("Critical", "15 minutes", 8, 7, "=D{}/C{}*100"),
        ("High", "1 hour", 35, 32, "=D{}/C{}*100"),
        ("Medium", "4 hours", 120, 115, "=D{}/C{}*100"),
        ("Low", "24 hours", 450, 448, "=D{}/C{}*100"),
    ]
    
    for severity, sla, incidents, within, formula in sla_examples:
        ws[f'A{row}'] = severity
        ws[f'B{row}'] = sla
        apply_style(ws[f'C{row}'], styles["input_cell"])
        apply_style(ws[f'D{row}'], styles["input_cell"])
        ws[f'C{row}'] = incidents
        ws[f'D{row}'] = within
        ws[f'E{row}'] = formula.format(row, row)
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_NARROW
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE
    
    # Freeze panes
    ws.freeze_panes = 'A4'

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 3
# ============================================================================

def create_soc_integration_sheet(wb, styles, validations):
    """Create SOC_Integration sheet."""
    ws = wb["SOC_Integration"]
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "SOC Integration Assessment"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 25
    
    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Assess SOC alert workflow, escalation, and staffing (20 criteria)"
    apply_style(ws['A2'], styles["subheader"])
    
    # Column headers
    headers = ["#", "Requirement", "Compliance", "Evidence ID", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Assessment criteria
    requirements = [
        "Are DLP alerts integrated with SOC workflow?",
        "Are Critical alerts delivered to SOC (SIEM + SMS/PagerDuty)?",
        "Are High alerts delivered to SOC (SIEM + Email)?",
        "Are SOC analysts trained on DLP alert triage?",
        "Is DLP training part of SOC analyst onboarding?",
        "Is DLP triage playbook documented (step-by-step)?",
        "Is escalation matrix documented (L1 → L2 → IR → Management)?",
        "Are SOC analysts able to escalate to DLP admin?",
        "Are SOC analysts able to escalate to user manager?",
        "Are SOC analysts able to escalate to Legal/DPO?",
        "Is SOC coverage 24/7 (for Critical DLP alerts)?",
        "Is SOC coverage business hours only (9-5) for Medium/Low alerts?",
        "Is on-call rotation defined (after-hours Critical alerts)?",
        "Are SOC metrics tracked (alert volume, triage time, SLA compliance)?",
        "Are SOC metrics reported to management (monthly)?",
        "Is SOC staffing adequate (not overwhelmed by alert volume)?",
        "Is SOC burnout monitored (high alert volume, turnover rate)?",
        "Are SOC tools adequate (SIEM, ticketing, communication)?",
        "Is SOC performance reviewed quarterly?",
        "Are SOC improvements implemented (based on performance review)?",
    ]
    
    row = 4
    # Add 3 example rows
    example_data = [
        (1, requirements[0], "Yes", "A812-4-SOC-001", "ServiceNow integration complete"),
        (2, requirements[3], "Partial", "A812-4-SOC-002", "Training scheduled for Q1"),
        (3, requirements[10], "Yes", "A812-4-SOC-003", "24/7 coverage with on-call"),
    ]
    
    for num, req, status, evidence, notes in example_data:
        ws[f'A{row}'] = num
        ws[f'B{row}'] = req
        ws[f'C{row}'] = status
        ws[f'D{row}'] = evidence
        ws[f'E{row}'] = notes
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["info_cell"])
        
        row += 1
    
    # Add remaining blank rows
    for idx, req in enumerate(requirements, start=1):
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = req
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["data_cell"])
        
        row += 1
    
    # Add data validation
    val_response = validations["response"]
    ws.add_data_validation(val_response)
    for r in range(7, 27):
        val_response.add(ws[f'C{r}'])
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_NARROW
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_dashboards_reporting_sheet(wb, styles, validations):
    """Create Dashboards_Reporting sheet."""
    ws = wb["Dashboards_Reporting"]
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "Dashboards & Reporting Assessment"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 25
    
    # Subheader
    ws.merge_cells('A2:E2')
    ws['A2'] = "Assess dashboard and reporting capabilities (20 criteria)"
    apply_style(ws['A2'], styles["subheader"])
    
    # Column headers
    headers = ["#", "Requirement", "Compliance", "Evidence ID", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Assessment criteria
    requirements = [
        "Is real-time SOC dashboard configured (alert queue)?",
        "SOC Dashboard: Shows Critical/High/Medium alerts pending triage?",
        "SOC Dashboard: Shows alert volume over time (hourly)?",
        "SOC Dashboard: Shows top blocked users (last 24 hours)?",
        "SOC Dashboard: Shows top blocked channels (email, USB, web)?",
        "SOC Dashboard: Shows false positive rate (rolling 7-day average)?",
        "Is executive dashboard configured (updated daily)?",
        "Executive Dashboard: Shows Critical incidents (last 30 days)?",
        "Executive Dashboard: Shows DLP effectiveness (prevented incidents)?",
        "Executive Dashboard: Shows compliance metrics (coverage %, enforcement rate)?",
        "Executive Dashboard: Shows risk heatmap (users, departments, channels)?",
        "Are weekly reports generated (Security Team)?",
        "Weekly Report: Alert summary (Critical/High/Medium/Low counts)?",
        "Weekly Report: Top users by alert volume?",
        "Weekly Report: False positive trend?",
        "Are monthly reports generated (Management)?",
        "Monthly Report: Executive summary (1-page, business language)?",
        "Monthly Report: Incident summary (critical events, resolved vs ongoing)?",
        "Are quarterly reports generated (CISO/Board)?",
        "Quarterly Report: Strategic risk assessment, ROI analysis?",
    ]
    
    row = 4
    # Add 3 example rows
    example_data = [
        (1, requirements[0], "Yes", "A812-4-DASH-001", "Splunk SOC dashboard live"),
        (2, requirements[6], "Partial", "A812-4-DASH-002", "Executive dashboard in development"),
        (3, requirements[11], "Yes", "A812-4-DASH-003", "Weekly reports automated"),
    ]
    
    for num, req, status, evidence, notes in example_data:
        ws[f'A{row}'] = num
        ws[f'B{row}'] = req
        ws[f'C{row}'] = status
        ws[f'D{row}'] = evidence
        ws[f'E{row}'] = notes
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["info_cell"])
        
        row += 1
    
    # Add remaining blank rows
    for idx, req in enumerate(requirements, start=1):
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = req
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles["data_cell"])
        
        row += 1
    
    # Add data validation
    val_response = validations["response"]
    ws.add_data_validation(val_response)
    for r in range(7, 27):
        val_response.add(ws[f'C{r}'])
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_NARROW
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_gap_analysis_sheet(wb, styles, validations):
    """Create Gap_Analysis sheet."""
    ws = wb["Gap_Analysis"]
    
    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "Gap Analysis - Monitoring & Response Deficiencies"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 25
    
    # Subheader
    ws.merge_cells('A2:J2')
    ws['A2'] = "Identify and track monitoring/response gaps and remediation plans (40 rows)"
    apply_style(ws['A2'], styles["subheader"])
    
    # Column headers
    headers = [
        "Gap ID", "Gap Description", "Affected Area", "Risk Level",
        "Business Impact", "Remediation Plan", "Owner", "Target Date",
        "Status", "Evidence ID"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Pre-populated examples (3 rows, gray)
    examples = [
        ("GAP-A812-4-001", "Critical alerts not reaching SOC SMS (delivery failures)", "Alerting",
         "High", "Delayed response to data exfiltration", "Configure PagerDuty integration",
         "SOC Lead", "2025-02-15", "In Progress", "A812-4-GAP-001"),
        ("GAP-A812-4-002", "SIEM correlation rules missing for insider threat use case", "SIEM",
         "Medium", "Cannot detect departing employee data theft", "Deploy insider threat correlation rule",
         "SIEM Admin", "2025-03-01", "Not Started", "A812-4-GAP-002"),
        ("GAP-A812-4-003", "False positive rate 12.5% for email DLP (above 10% target)", "FP Management",
         "Medium", "Alert fatigue, SOC overwhelmed", "Tune email DLP rules, whitelist legitimate flows",
         "DLP Admin", "2025-02-28", "In Progress", "A812-4-GAP-003"),
    ]
    
    row = 4
    for gap_data in examples:
        for col_idx, value in enumerate(gap_data, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # Add blank rows (37 more = 40 total)
    for gap_num in range(4, 41):
        ws[f'A{row}'] = f'GAP-A812-4-{gap_num:03d}'
        for col_idx in range(2, 11):
            apply_style(ws.cell(row=row, column=col_idx), styles["data_cell"])
        row += 1
    
    # Add data validations
    val_affected_area = validations["affected_area"]
    val_risk = validations["risk"]
    val_gap_status = validations["gap_status"]
    
    ws.add_data_validation(val_affected_area)
    ws.add_data_validation(val_risk)
    ws.add_data_validation(val_gap_status)
    
    for r in range(7, 44):  # Rows 7-43 (after 3 examples)
        val_affected_area.add(ws[f'C{r}'])
        val_risk.add(ws[f'D{r}'])
        val_gap_status.add(ws[f'I{r}'])
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['F'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['G'].width = WIDTH_WIDE
    ws.column_dimensions['H'].width = WIDTH_MEDIUM
    ws.column_dimensions['I'].width = WIDTH_MEDIUM
    ws.column_dimensions['J'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A4'


def create_evidence_register_sheet(wb, styles, validations):
    """Create Evidence_Register sheet."""
    ws = wb["Evidence_Register"]
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = "Evidence Register - Domain 4 (Monitoring & Response)"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 25
    
    # Subheader
    ws.merge_cells('A2:H2')
    ws['A2'] = "Track all evidence collected for monitoring and response assessment (100 rows)"
    apply_style(ws['A2'], styles["subheader"])
    
    # Column headers
    headers = [
        "Evidence ID", "Evidence Type", "Description", "Location/Link",
        "Date Collected", "Collected By", "Related Requirement", "Verification Status"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Pre-populated examples (5 rows, gray)
    examples = [
        ("A812-4-LOG-001", "Screenshot", "SIEM showing DLP log ingestion", "/evidence/siem_dlp_logs.png",
         "2025-01-06", "SIEM Admin", "Logging_Configuration #1", "Verified"),
        ("A812-4-ALT-001", "Config", "Alert rule configuration export", "/evidence/alert_rules.json",
         "2025-01-06", "DLP Admin", "Alert_Rules_Inventory", "Verified"),
        ("A812-4-SIEM-001", "Screenshot", "SIEM correlation rule for insider threat", "/evidence/correlation_rule.png",
         "2025-01-06", "SOC Lead", "SIEM_Integration #12", "Verified"),
        ("A812-4-FP-001", "Dashboard", "False positive rate dashboard", "/evidence/fp_dashboard.png",
         "2025-01-06", "DLP Admin", "False_Positive_Management #1", "Verified"),
        ("A812-4-IR-001", "Report", "Incident response SLA compliance report", "/evidence/sla_report.pdf",
         "2025-01-06", "SOC Lead", "Incident_Response_Workflow #23", "Verified"),
    ]
    
    row = 4
    for evidence_data in examples:
        for col_idx, value in enumerate(evidence_data, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # Add blank rows (95 more = 100 total)
    for _ in range(95):
        for col_idx in range(1, 9):
            apply_style(ws.cell(row=row, column=col_idx), styles["data_cell"])
        row += 1
    
    # Add data validations
    val_evidence = validations["evidence"]
    val_verification = validations["verification"]
    
    ws.add_data_validation(val_evidence)
    ws.add_data_validation(val_verification)
    
    for r in range(9, 104):  # Rows 9-103 (after 5 examples)
        val_evidence.add(ws[f'B{r}'])
        val_verification.add(ws[f'H{r}'])
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_DESCRIPTION
    ws.column_dimensions['D'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['E'].width = WIDTH_MEDIUM
    ws.column_dimensions['F'].width = WIDTH_MEDIUM
    ws.column_dimensions['G'].width = WIDTH_WIDE
    ws.column_dimensions['H'].width = WIDTH_MEDIUM
    
    # Freeze panes
    ws.freeze_panes = 'A4'

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 4
# ============================================================================

def create_summary_dashboard_sheet(wb, styles):
    """Create Summary_Dashboard sheet."""
    ws = wb["Summary_Dashboard"]
    
    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = "Domain 4 - Monitoring & Response Summary Dashboard"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Metadata section
    row = 3
    ws[f'A{row}'] = "Assessment Date:"
    ws[f'A{row}'].font = Font(bold=True)
    apply_style(ws[f'B{row}'], styles["input_cell"])
    
    row += 1
    ws[f'A{row}'] = "Completed By:"
    ws[f'A{row}'].font = Font(bold=True)
    apply_style(ws[f'B{row}'], styles["input_cell"])
    
    row += 1
    ws[f'A{row}'] = "Approved By:"
    ws[f'A{row}'].font = Font(bold=True)
    apply_style(ws[f'B{row}'], styles["input_cell"])
    
    row += 1
    ws[f'A{row}'] = "Next Review:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = '=B3+90'
    ws[f'B{row}'].number_format = 'DD.MM.YYYY'
    
    row += 2
    
    # KPI Section
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "OVERALL DOMAIN 4 COMPLIANCE SCORE"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    
    # Column headers for KPIs
    kpi_headers = ["Metric", "Value", "Target", "Status"]
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    row += 1
    
    # KPIs with formulas
    kpis = [
        ("Logging Configuration Complete %", 
         '=COUNTIF(Logging_Configuration!C:C,"Yes")/25*100', "90%"),
        ("Alert Rules Documented %",
         '=COUNTA(Alert_Rules_Inventory!A4:A33)/30*100', "80%"),
        ("SIEM Integration Complete %",
         '=COUNTIF(SIEM_Integration!C:C,"Yes")/25*100', "90%"),
        ("False Positive Rate %",
         '=AVERAGE(Alert_Rules_Inventory!J4:J33)', "<10%"),
        ("Incident Response SLA Compliance %",
         '=AVERAGE(Incident_Response_Workflow!E35:E38)', ">90%"),
        ("SOC Integration Complete %",
         '=COUNTIF(SOC_Integration!C:C,"Yes")/20*100', "85%"),
        ("Dashboards Deployed %",
         '=COUNTIF(Dashboards_Reporting!C:C,"Yes")/20*100', "80%"),
    ]
    
    kpi_start_row = row
    for kpi_name, formula, target in kpis:
        ws[f'A{row}'] = kpi_name
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = '0.0'
        ws[f'C{row}'] = target
        
        # Status formula (traffic light)
        if "Rate" in kpi_name and "False" in kpi_name:
            # FP rate: lower is better
            ws[f'D{row}'] = f'=IF(B{row}<=10,"\u2705 On Target",IF(B{row}<=15,"\u26A0\uFE0F Close","\u274C Action Needed"))'
        elif "SLA" in kpi_name:
            # SLA: higher is better
            ws[f'D{row}'] = f'=IF(B{row}>=90,"\u2705 On Target",IF(B{row}>=80,"\u26A0\uFE0F Close","\u274C Action Needed"))'
        else:
            # Standard percentage: higher is better
            target_num = float(target.replace('%', '').replace('>', ''))
            ws[f'D{row}'] = f'=IF(B{row}>={target_num},"\u2705 On Target",IF(B{row}>={target_num-10},"\u26A0\uFE0F Close","\u274C Action Needed"))'
        
        row += 1
    
    # Overall compliance score
    ws[f'A{row}'] = "OVERALL DOMAIN 4 COMPLIANCE"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'] = f'=AVERAGE(B{kpi_start_row}:B{row-1})'
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].number_format = '0.0'
    ws[f'C{row}'] = ">85%"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f'=IF(B{row}>=85,"\u2705 Pass",IF(B{row}>=70,"\u26A0\uFE0F Conditional","\u274C Fail"))'
    ws[f'D{row}'].font = Font(bold=True, size=12)
    
    row += 2
    
    # Key Findings Section
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "KEY FINDINGS"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    
    findings = [
        ("Total Alerts per Day (Avg):", '=Alert_Volume_Metrics!C7'),
        ("False Positive Rate (Avg):", '=AVERAGE(Alert_Rules_Inventory!J4:J33)'),
        ("MTTD (Mean Time to Detect):", ""),
        ("MTTR (Mean Time to Respond):", ""),
        ("Critical Gaps Count:", '=COUNTIF(Gap_Analysis!D:D,"Critical")'),
        ("Evidence Completeness %:", '=COUNTA(Evidence_Register!A4:A103)/100*100'),
    ]
    
    for finding_name, formula in findings:
        ws[f'A{row}'] = finding_name
        ws[f'A{row}'].font = Font(bold=True)
        if formula:
            ws[f'B{row}'] = formula
            if "%" in finding_name:
                ws[f'B{row}'].number_format = '0.0'
        else:
            apply_style(ws[f'B{row}'], styles["input_cell"])
        row += 1
    
    row += 1
    
    # Approval Workflow Section
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "APPROVAL WORKFLOW"
    apply_style(ws[f'A{row}'], styles["subheader"])
    
    row += 1
    
    approval_headers = ["Role", "Name", "Signature", "Date"]
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    row += 1
    
    approvers = ["DLP Administrator", "SOC Lead", "CISO", "DPO"]
    for approver in approvers:
        ws[f'A{row}'] = approver
        for col in ['B', 'C', 'D']:
            apply_style(ws[f'{col}{row}'], styles["input_cell"])
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_WIDE
    
    # Freeze panes
    ws.freeze_panes = 'A3'


# ============================================================================
# SECTION 5: MAIN FUNCTION
# ============================================================================

def main():
    """Main function to generate the workbook."""
    print(f"Generating {WORKBOOK_ID} - {ASSESSMENT_AREA}...")
    
    # Create workbook
    wb = create_workbook()
    
    # Setup styles
    styles = setup_styles()
    
    # Setup data validations
    validations = setup_data_validations(wb)
    
    # Create all sheets
    print("Creating Instructions_Legend...")
    create_instructions_sheet(wb, styles)
    
    print("Creating Logging_Configuration...")
    create_logging_configuration_sheet(wb, styles, validations)
    
    print("Creating Alert_Rules_Inventory...")
    create_alert_rules_inventory_sheet(wb, styles, validations)
    
    print("Creating Alert_Volume_Metrics...")
    create_alert_volume_metrics_sheet(wb, styles)
    
    print("Creating SIEM_Integration...")
    create_siem_integration_sheet(wb, styles, validations)
    
    print("Creating False_Positive_Management...")
    create_false_positive_management_sheet(wb, styles, validations)
    
    print("Creating Incident_Response_Workflow...")
    create_incident_response_workflow_sheet(wb, styles, validations)
    
    print("Creating SOC_Integration...")
    create_soc_integration_sheet(wb, styles, validations)
    
    print("Creating Dashboards_Reporting...")
    create_dashboards_reporting_sheet(wb, styles, validations)
    
    print("Creating Gap_Analysis...")
    create_gap_analysis_sheet(wb, styles, validations)
    
    print("Creating Evidence_Register...")
    create_evidence_register_sheet(wb, styles, validations)
    
    print("Creating Summary_Dashboard...")
    create_summary_dashboard_sheet(wb, styles)
    
    # Save workbook
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.8.12.4_Monitoring_Response_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    wb.save(filename)
    print(f"\n\u2705 Success! Workbook saved as: {filename}")
    print(f"\nWorkbook Statistics:")
    print(f"  - Total Sheets: 12")
    print(f"  - Assessment Items: ~70 monitoring/response criteria")
    print(f"  - Alert Rules: 30 rows")
    print(f"  - Gap Analysis: 40 rows")
    print(f"  - Evidence Register: 100 rows")
    print(f"\n🎯 Ready for Domain 4 (Monitoring & Response) assessment!")


if __name__ == "__main__":
    main()