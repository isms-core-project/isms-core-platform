#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS A.8.12 - DLP Dashboard Data Consolidation Script
================================================================================

ISO/IEC 27001:2022 Annex A Control A.8.12: Data Leakage Prevention
Dashboard Utility: Static Data Consolidation

--------------------------------------------------------------------------------
PURPOSE
--------------------------------------------------------------------------------

Consolidates user-entered assessment data from 4 DLP workbooks into executive
compliance dashboard for CISO reporting and audit evidence packages, creating
point-in-time compliance snapshot.

This utility provides ALTERNATIVE to external workbook formulas by directly
copying data values (not formulas) from source assessments into dashboard detail
sheets, useful for archival, distribution, and situations where external formulas
cannot be used.

When to use:
• Creating quarterly audit evidence packages (static snapshots)
• Distributing dashboard without source files (single-file delivery)
• Excel security policies prevent external workbook links
• Historical compliance tracking (dated snapshots)

When NOT to use:
• Day-to-day compliance monitoring (use external formulas instead)
• Real-time gap tracking (external formulas update automatically)

--------------------------------------------------------------------------------
WHAT IT DOES
--------------------------------------------------------------------------------

1. Validates all 4 source assessment workbooks exist (normalized filenames)
2. Extracts gaps from each workbook (Partial/Non-Compliant items)
3. Extracts evidence entries from Evidence_Register sheets
4. Auto-generates risks from High severity gaps
5. Auto-generates remediation actions for all gaps
6. Consolidates data into dashboard detail sheets:
   • Consolidated_Gap_Analysis (all gaps from domains 1-4)
   • Risk_Register (DLP-related risks)
   • Remediation_Roadmap (action plans with timelines)
   • Evidence_Master_Index (consolidated evidence)
7. Preserves Executive_Summary sheet (does not modify formulas)
8. Generates consolidation statistics report

Data Extraction Logic:
    Gaps: Scans assessment sheets for [!] Partial or [X] Non-Compliant status
    Risks: Creates risk entry for each High severity gap
    Remediation: Creates action for each gap with calculated target dates:
        • Critical gaps: 30 days
        • High gaps: 60 days
        • Medium gaps: 90 days
    Evidence: Reads all Evidence_Register entries from each workbook

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System:
    Python 3.8 or higher

Dependencies:
    openpyxl - Excel file reading and writing
    
Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic:
    python3 consolidate_a812_dashboard.py ISMS-IMP-A.8.12.5_Dashboard_YYYYMMDD.xlsx

Arguments:
    dashboard_file: Path to generated dashboard workbook

Prerequisites:
    1. All 4 assessment workbooks completed (user data entered)
    2. Files normalized: python3 normalize_assessment_files_a812.py
    3. Dashboard generated: python3 generate_a812_5_compliance_dashboard.py
    4. All 5 files in current directory (same location)

Required Source Files (in current directory):
    • ISMS-IMP-A.8.12.1.xlsx (DLP Infrastructure Assessment)
    • ISMS-IMP-A.8.12.2.xlsx (Data Classification Assessment)
    • ISMS-IMP-A.8.12.3.xlsx (Channel Coverage Assessment)
    • ISMS-IMP-A.8.12.4.xlsx (Monitoring & Response Assessment)
    • [Dashboard file specified as argument]

Output:
    Modified dashboard with consolidated data in detail sheets
    Console statistics showing consolidation results

Examples:
    # Standard usage
    python3 consolidate_a812_dashboard.py ISMS-IMP-A.8.12.5_Dashboard_20250125.xlsx
    
    # Create quarterly snapshot
    python3 consolidate_a812_dashboard.py ISMS-IMP-A.8.12.5_Q1_2025_Audit_Package.xlsx
    
    # Historical compliance tracking
    python3 consolidate_a812_dashboard.py ISMS-IMP-A.8.12.5_Dashboard_20250125.xlsx
    cp ISMS-IMP-A.8.12.5_Dashboard_20250125.xlsx ../Archive/Q1_2025_Snapshot.xlsx

Post-Consolidation:
    • Open dashboard - no "Update Links" prompt (static data)
    • Review consolidation statistics for data volume
    • Verify gap, risk, and evidence counts match expectations
    • Dashboard ready for distribution (no dependencies)

--------------------------------------------------------------------------------
WORKFLOW INTEGRATION
--------------------------------------------------------------------------------

Position in A.8.12 Framework:
    Optional alternative to external workbook formulas
    Creates static point-in-time snapshots for archival and distribution

Typical Usage Pattern:
    Use for quarterly audit packages, Board presentations, or when external
    formulas cannot be used due to Excel security policies or distribution
    requirements.

Integration Workflow:
    1. Generate assessment workbooks (scripts 1-4)
    2. Complete assessments (manual work)
    3. Normalize filenames (normalize_assessment_files_a812.py)
    4. Generate dashboard (generate_a812_5_compliance_dashboard.py)
    5. Run THIS SCRIPT (consolidate data)              ← YOU ARE HERE
    6. Dashboard now contains static snapshot
    7. Distribute to CISO/auditors (single file)

Related Scripts:
    - generate_a812_5_compliance_dashboard.py (creates dashboard with formulas)
    - normalize_assessment_files_a812.py (prerequisite - creates normalized files)
    - excel_sanity_check_a812.py (validates workbooks before consolidation)

Comparison with External Formulas:
    External Formulas (Dashboard Default):
        ✅ Real-time updates when source data changes
        ✅ No manual refresh required
        ❌ Requires files co-located
        ❌ Excel security warnings
        ❌ Cannot distribute dashboard alone
    
    Data Consolidation (This Script):
        ✅ Single-file distribution (no dependencies)
        ✅ No Excel security warnings
        ✅ Historical snapshots for audit trail
        ❌ Manual refresh required (re-run script)
        ❌ Larger file size (data duplicated)
        ❌ Not real-time (static snapshot)

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.12
Utility Type:         Data Consolidation / Static Snapshot Creation
Framework Version:    1.0
Script Version:       1.0
Date:                 25.01.2025
Author:               [Organization] ISMS Implementation Team

--------------------------------------------------------------------------------
NOTES
--------------------------------------------------------------------------------

Re-Running Consolidation:
    Safe to re-run when source assessment data updated.
    Existing consolidated data will be overwritten.
    Statistics show current consolidation results only (not cumulative).

Source File Preservation:
    Source assessment workbooks opened in read-only mode.
    No modifications made to source files.
    Only dashboard workbook is modified (detail sheets populated).

Executive Summary Preservation:
    Executive_Summary sheet is NOT modified by this script.
    Existing external formulas in Executive_Summary remain intact.
    This allows hybrid approach: formulas for summary + consolidated detail data.

Gap ID Format:
    GAP-1-001 (Domain 1, Gap #1)
    GAP-2-015 (Domain 2, Gap #15)
    GAP-3-007 (Domain 3, Gap #7)
    GAP-4-042 (Domain 4, Gap #42)

Target Date Calculation:
    Remediation target dates calculated from today:
    • Critical severity: Today + 30 days
    • High severity: Today + 60 days
    • Medium severity: Today + 90 days
    • Low severity: Today + 120 days

Data Traceability:
    Each consolidated entry includes source domain label:
    • "Domain 1 - Infrastructure"
    • "Domain 2 - Classification"
    • "Domain 3 - Channels"
    • "Domain 4 - Monitoring"
    
    Enables traceability back to source assessment workbook.

Audit Considerations:
    Consolidated dashboards serve as point-in-time audit evidence.
    Recommended practice:
    • Create quarterly consolidated snapshots
    • Date filename (Q1_2025, Q2_2025, etc.)
    • Archive with source assessment files
    • Maintain 3 years of snapshots for audit cycle

Performance:
    Typical consolidation time: 10-30 seconds
    Depends on data volume in source assessments
    Progress displayed for each source workbook
    Large datasets (1000+ gaps) may take longer

================================================================================
END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from datetime import datetime, timedelta

# ============================================================================
# CONFIGURATION
# ============================================================================

SOURCE_WORKBOOKS = {
    "wb1": "ISMS-IMP-A.8.12.1.xlsx",  # DLP Infrastructure
    "wb2": "ISMS-IMP-A.8.12.2.xlsx",  # Data Classification
    "wb3": "ISMS-IMP-A.8.12.3.xlsx",  # Channel Coverage
    "wb4": "ISMS-IMP-A.8.12.4.xlsx",  # Monitoring & Response
}

DOMAIN_NAMES = {
    "wb1": "Domain 1 - Infrastructure",
    "wb2": "Domain 2 - Classification",
    "wb3": "Domain 3 - Channels",
    "wb4": "Domain 4 - Monitoring",
}

# Sheet starting rows in dashboard
SHEET_START_ROWS = {
    "Consolidated_Gap_Analysis": 6,
    "Risk_Register": 9,  # Skip example risks (rows 4-8)
    "Remediation_Roadmap": 4,
    "Evidence_Master_Index": 4,
}

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def safely_write_data(ws, start_row, data):
    """Safely write data, handling merged cells"""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except:
                continue
        entries_written += 1
    return entries_written

# ============================================================================
# DATA EXTRACTION FUNCTIONS
# ============================================================================

def extract_gaps_from_workbook(filepath, domain_name):
    """Extract gap items (Partial/Non-Compliant) from source workbook"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        gaps = []
        gap_counter = 1
        
        # Iterate through assessment sheets
        for sheet_name in wb.sheetnames:
            # Skip non-assessment sheets
            if sheet_name in ['Instructions', 'Summary_Dashboard', 'Evidence_Register', 
                             'Approval_Sign_Off', 'Document_Control']:
                continue
            
            ws = wb[sheet_name]
            
            # Find status column and gap description column
            # Typically: Status in column with [OK]/[!]/[X], Gap nearby
            for row in range(8, min(ws.max_row + 1, 100)):
                row_values = [ws.cell(row=row, column=col).value for col in range(1, 20)]
                
                # Find status (look for [!] or [X])
                status = None
                status_col = None
                for col_idx, val in enumerate(row_values, start=1):
                    if val in ['[!] Partial', '[X] Non-Compliant']:
                        status = val
                        status_col = col_idx
                        break
                
                if not status:
                    continue
                
                # Extract data for this gap
                system_name = row_values[0] if row_values[0] else 'Unknown'
                gap_desc = None
                
                # Gap description typically near status column
                if status_col:
                    # Try to find gap description (usually 2-3 columns after status)
                    for offset in [2, 3, 4]:
                        potential_gap = ws.cell(row=row, column=status_col + offset).value
                        if potential_gap and len(str(potential_gap)) > 10:
                            gap_desc = potential_gap
                            break
                
                severity = 'High' if status == '[X] Non-Compliant' else 'Medium'
                
                gap_entry = [
                    domain_name,                                          # Domain
                    f'GAP-{domain_name.split()[1]}-{gap_counter:03d}',  # Gap ID
                    gap_desc if gap_desc else f'{system_name}: Configuration gap identified',  # Gap Description
                    severity,                                             # Severity
                    status.replace('[!] ', '').replace('[X] ', ''),      # Current State
                    'Compliant',                                          # Target State
                    None,                                                 # Owner (to be assigned)
                    'Open',                                               # Status
                ]
                gaps.append(gap_entry)
                gap_counter += 1
        
        wb.close()
        return gaps
    except Exception as e:
        print(f"  [!] Error extracting gaps from {filepath}: {e}")
        return []

def extract_evidence_from_workbook(filepath, domain_name):
    """Extract evidence entries from source workbook"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        if 'Evidence_Register' not in wb.sheetnames:
            wb.close()
            return []
        
        ws = wb['Evidence_Register']
        evidence = []
        
        for row in range(10, ws.max_row + 1):
            evidence_id = ws[f'A{row}'].value
            if evidence_id and str(evidence_id).startswith('EVD-'):
                row_data = [
                    evidence_id,                      # Evidence ID
                    domain_name,                      # Domain
                    ws[f'B{row}'].value,             # Control/Requirement
                    ws[f'C{row}'].value,             # Evidence Type
                    ws[f'D{row}'].value,             # File Location
                    ws[f'E{row}'].value,             # Collected Date
                    ws[f'F{row}'].value,             # Verified By
                    ws[f'G{row}'].value,             # Status
                ]
                evidence.append(row_data)
        
        wb.close()
        return evidence
    except Exception as e:
        print(f"  [!] Error reading evidence from {filepath}: {e}")
        return []

# ============================================================================
# DATA GENERATION FUNCTIONS
# ============================================================================

def generate_risks_from_gaps(gaps):
    """Generate risk entries from critical gaps"""
    risks = []
    risk_counter = 1
    
    for gap in gaps:
        domain = gap[0]
        gap_id = gap[1]
        gap_desc = gap[2]
        severity = gap[3]
        
        # Only create risks for High severity gaps
        if severity != 'High':
            continue
        
        risk_entry = [
            f'RISK-DLP-{risk_counter:03d}',                      # Risk ID
            domain,                                               # Domain
            f'Security risk: {gap_desc[:80]}...',                # Risk Description
            'Technical',                                          # Category
            'High',                                               # Likelihood
            'High',                                               # Impact
            'High-High',                                          # Risk Score
            'High',                                               # Inherent Risk
            'Medium',                                             # Residual Risk (target)
            'Active',                                             # Status
        ]
        risks.append(risk_entry)
        risk_counter += 1
    
    return risks

def generate_remediation_from_gaps(gaps):
    """Generate remediation roadmap from gaps"""
    remediation = []
    action_counter = 1
    today = datetime.now()
    
    for gap in gaps:
        domain = gap[0]
        gap_id = gap[1]
        gap_desc = gap[2]
        severity = gap[3]
        
        # Priority based on severity
        priority = 'Critical' if severity == 'High' else 'High' if severity == 'Medium' else 'Medium'
        
        # Target date based on priority
        if priority == 'Critical':
            days_offset = 30
        elif priority == 'High':
            days_offset = 60
        else:
            days_offset = 90
        
        target_date = (today + timedelta(days=days_offset)).strftime('%Y-%m-%d')
        
        remediation_entry = [
            f'REM-DLP-{action_counter:03d}',                     # Action ID
            domain,                                               # Domain
            f'Remediate: {gap_desc[:60]}...',                    # Action Description
            priority,                                             # Priority
            None,                                                 # Owner
            today.strftime('%Y-%m-%d'),                          # Start Date
            target_date,                                          # Target Date
            None,                                                 # Actual Date
            None,                                                 # Budget
            'Planned',                                            # Status
        ]
        remediation.append(remediation_entry)
        action_counter += 1
    
    return remediation

# ============================================================================
# MAIN CONSOLIDATION FUNCTION
# ============================================================================

def populate_comprehensive_dashboard(dashboard_file):
    """Populate all dashboard sheets from source workbooks"""
    
    print("="*80)
    print("A.8.12 DLP COMPREHENSIVE DASHBOARD POPULATION")
    print("="*80)
    
    # Check for source workbooks
    missing_workbooks = []
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        if not os.path.exists(wb_filename):
            missing_workbooks.append(wb_filename)
    
    if missing_workbooks:
        print("\n[X] ERROR: Missing source workbooks:")
        for wb in missing_workbooks:
            print(f"    - {wb}")
        print("\n  Place all 4 assessment workbooks in the same directory as this script.")
        return False
    
    all_gaps = []
    all_evidence = []
    
    print("\n[1/4] Extracting gaps from source workbooks...")
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        domain_name = DOMAIN_NAMES[wb_key]
        print(f"  * {domain_name}: ", end="")
        gaps = extract_gaps_from_workbook(wb_filename, domain_name)
        print(f"{len(gaps)} gaps")
        all_gaps.extend(gaps)
    
    print(f"\n  Total gaps identified: {len(all_gaps)}")
    
    print("\n[2/4] Extracting evidence from source workbooks...")
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        domain_name = DOMAIN_NAMES[wb_key]
        print(f"  * {domain_name}: ", end="")
        evidence = extract_evidence_from_workbook(wb_filename, domain_name)
        print(f"{len(evidence)} evidence docs")
        all_evidence.extend(evidence)
    
    print(f"\n  Total evidence collected: {len(all_evidence)}")
    
    print("\n[3/4] Generating risks and remediation...")
    all_risks = generate_risks_from_gaps(all_gaps)
    all_remediation = generate_remediation_from_gaps(all_gaps)
    print(f"  * Risks generated: {len(all_risks)}")
    print(f"  * Remediation actions: {len(all_remediation)}")
    
    print(f"\n[4/4] Writing to dashboard...")
    try:
        wb = load_workbook(dashboard_file)
    except Exception as e:
        print(f"  [X] Error loading dashboard: {e}")
        return False
    
    # Populate Consolidated Gap Analysis
    if 'Consolidated_Gap_Analysis' in wb.sheetnames and all_gaps:
        ws = wb['Consolidated_Gap_Analysis']
        count = safely_write_data(ws, SHEET_START_ROWS['Consolidated_Gap_Analysis'], all_gaps)
        print(f"  [OK] Consolidated_Gap_Analysis: {count} entries")
    
    # Populate Risk Register
    if 'Risk_Register' in wb.sheetnames and all_risks:
        ws = wb['Risk_Register']
        count = safely_write_data(ws, SHEET_START_ROWS['Risk_Register'], all_risks)
        print(f"  [OK] Risk_Register: {count} entries")
    
    # Populate Remediation Roadmap
    if 'Remediation_Roadmap' in wb.sheetnames and all_remediation:
        ws = wb['Remediation_Roadmap']
        count = safely_write_data(ws, SHEET_START_ROWS['Remediation_Roadmap'], all_remediation)
        print(f"  [OK] Remediation_Roadmap: {count} entries")
    
    # Populate Evidence Master Index
    if 'Evidence_Master_Index' in wb.sheetnames and all_evidence:
        ws = wb['Evidence_Master_Index']
        count = safely_write_data(ws, SHEET_START_ROWS['Evidence_Master_Index'], all_evidence)
        print(f"  [OK] Evidence_Master_Index: {count} entries")
    
    # Save
    try:
        wb.save(dashboard_file)
        print(f"\n[BACKUP] Saved: {dashboard_file}")
    except Exception as e:
        print(f"  [X] Error saving: {e}")
        return False
    
    print("\n" + "="*80)
    print("[OK] A.8.12 DLP DASHBOARD POPULATION COMPLETE")
    print("="*80)
    print(f"\nSummary:")
    print(f"  * Consolidated Gap Analysis: {len(all_gaps)} gaps identified")
    print(f"  * Risk Register: {len(all_risks)} risks documented")
    print(f"  * Remediation Roadmap: {len(all_remediation)} actions planned")
    print(f"  * Evidence Master Index: {len(all_evidence)} evidence documents")
    print(f"\n  Total Data Points: {len(all_gaps) + len(all_risks) + len(all_remediation) + len(all_evidence)}")
    print("\n[ROCKET] Dashboard is now CISO-presentation ready!")
    print("="*80 + "\n")
    
    return True

# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 consolidate_a812_dashboard.py <dashboard.xlsx>")
        print("\nExample:")
        print("  python3 consolidate_a812_dashboard.py ISMS-IMP-A.8.12.5_Compliance_Dashboard_20260117.xlsx")
        sys.exit(1)
    
    dashboard_file = sys.argv[1]
    
    if not os.path.exists(dashboard_file):
        print(f"[X] Error: Dashboard file not found: {dashboard_file}")
        sys.exit(1)
    
    success = populate_comprehensive_dashboard(dashboard_file)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
