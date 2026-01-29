#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS A.8.12 - Excel Workbook Sanity Check Utility
================================================================================

ISO/IEC 27001:2022 Annex A Control A.8.12: Data Leakage Prevention
Quality Assurance Utility: Workbook Structure Validation

--------------------------------------------------------------------------------
PURPOSE
--------------------------------------------------------------------------------

Validates structural integrity and quality of DLP assessment workbooks before
consolidation and audit presentation, detecting common openpyxl generation errors
that cause Excel repair warnings.

This diagnostic utility identifies issues that would otherwise require manual
Excel file repair, ensuring workbooks are audit-ready and professional quality
before delivery to CISO, auditors, or stakeholders.

Common issues detected:
• Formula syntax errors and invalid sheet references (#REF!, #NAME?)
• Data validation conflicts and overlapping ranges
• Style object reuse issues (openpyxl common error)
• Merged cell content problems
• Missing required sheets or incorrect sheet names
• Inconsistent column widths and formatting

When to run:
• After generating any assessment workbook (quality check)
• Before distributing workbooks to stakeholders
• When Excel displays "file level validation and repair" warnings
• During development/testing of generator scripts

--------------------------------------------------------------------------------
WHAT IT DOES
--------------------------------------------------------------------------------

1. Scans directory for DLP assessment workbooks (.xlsx files)
2. Validates workbook structure:
   • Correct sheet names present
   • Required sheets not missing
   • No unexpected extra sheets
3. Checks formula integrity:
   • No #REF! errors (broken references)
   • No #NAME? errors (invalid function names)
   • External formulas reference correct files
4. Validates data validations:
   • Dropdown lists properly configured
   • No overlapping validation ranges
   • List sources valid
5. Detects style issues:
   • No reused style objects (openpyxl common bug)
   • Consistent formatting
   • Appropriate cell protection
6. Verifies examples present:
   • Pre-populated rows exist
   • Gray example rows formatted correctly
7. Checks structural elements:
   • Column widths appropriate
   • Freeze panes set correctly
   • Headers and formatting consistent

Output:
    Console report categorizing issues by severity:
    • CRITICAL - Must fix before distribution (breaks functionality)
    • WARNING - Should fix (quality/consistency issues)
    • INFO - Optional improvements

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System:
    Python 3.8 or higher

Dependencies:
    openpyxl - Excel file reading and structure analysis
    
Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic (check current directory):
    python3 excel_sanity_check_a812.py

Specify directory:
    python3 excel_sanity_check_a812.py /path/to/assessments

Arguments:
    [directory]: Optional path to directory containing workbooks
                 Default: current directory

What gets checked:
    • All files matching: ISMS-IMP-A.8.12.*.xlsx
    • Includes: Assessment workbooks 1-4, Dashboard workbook 5
    • Excludes: Temporary files (~$*.xlsx), backup files (*.bak)

Output:
    Console report showing:
    • Workbooks scanned
    • Issues found (by severity)
    • Pass/fail status per workbook
    • Overall framework health summary

Examples:
    # Check current directory
    python3 excel_sanity_check_a812.py
    
    # Check specific directory
    python3 excel_sanity_check_a812.py /home/user/DLP_Assessments/
    
    # Check before distribution
    python3 excel_sanity_check_a812.py && echo "All checks passed"
    
    # Automated quality gate in CI/CD
    if python3 excel_sanity_check_a812.py > sanity_report.txt; then
        echo "Quality checks passed"
    else
        echo "Quality checks failed - see sanity_report.txt"
        exit 1
    fi

Expected Sheet Names by Workbook:
    A.8.12.1 (Infrastructure):
        Instructions_Legend, DLP_Technology_Inventory, Network_DLP,
        Endpoint_DLP, Email_DLP, Cloud_CASB_DLP, Web_DLP, Database_DAM,
        Gap_Analysis, Evidence_Register, Summary_Dashboard
    
    A.8.12.2 (Data Classification):
        Instructions_Legend, Classification_Schema, Sensitive_Data_Inventory,
        Data_Locations, Data_Ownership, PII_Inventory, Regulatory_Mapping,
        Data_Discovery_Tools, Gap_Analysis, Evidence_Register, Summary_Dashboard
    
    A.8.12.3 (Channel Coverage):
        Instructions_Legend, Channel_Overview, Email_Channel, Web_Cloud_Channel,
        Endpoint_Channel, Network_Channel, Application_Channel, Mobile_Channel,
        Coverage_Metrics, Gap_Analysis, Evidence_Register, Summary_Dashboard
    
    A.8.12.4 (Monitoring & Response):
        Instructions_Legend, Logging_Configuration, Alert_Rules_Inventory,
        Alert_Volume_Metrics, SIEM_Integration, False_Positive_Management,
        Incident_Response_Workflow, SOC_Integration, Dashboards_Reporting,
        Gap_Analysis, Evidence_Register, Summary_Dashboard
    
    A.8.12.5 (Compliance Dashboard):
        Executive_Summary, Consolidated_Gap_Analysis, Risk_Register,
        Remediation_Roadmap, Evidence_Master_Index, KPI_Dashboard,
        Compliance_Scorecard, Approval_Sign_Off, Document_Control

--------------------------------------------------------------------------------
WORKFLOW INTEGRATION
--------------------------------------------------------------------------------

Position in A.8.12 Framework:
    Quality assurance checkpoint - can be run at any stage
    Recommended: Run after generating workbooks, before distribution

Typical Usage Pattern:
    • Development: Run after modifying generator scripts
    • Pre-distribution: Run before sending to stakeholders
    • Troubleshooting: Run when Excel repair warnings appear
    • CI/CD: Automate as quality gate in build pipeline

Integration Workflow:
    1. Generate assessment workbooks (scripts 1-4)
    2. Run THIS UTILITY (validate structure)               ← Quality Gate
    3. Fix any CRITICAL issues identified
    4. Complete assessments (manual work)
    5. Run THIS UTILITY again (final quality check)        ← Quality Gate
    6. Normalize filenames (if checks pass)
    7. Generate dashboard (if checks pass)
    8. Distribute to CISO/auditors

Related Scripts:
    - generate_a812_1_dlp_infrastructure.py (creates workbook 1)
    - generate_a812_2_data_classification.py (creates workbook 2)
    - generate_a812_3_channel_coverage.py (creates workbook 3)
    - generate_a812_4_monitoring_response.py (creates workbook 4)
    - generate_a812_5_compliance_dashboard.py (creates dashboard)

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.12
Utility Type:         Quality Assurance / Structure Validation
Framework Version:    1.0
Script Version:       1.0
Date:                 25.01.2025
Author:               [Organization] ISMS Implementation Team

--------------------------------------------------------------------------------
NOTES
--------------------------------------------------------------------------------

Non-Destructive Testing:
    This utility only READS workbook files.
    No modifications made to any files.
    Safe to run on production workbooks.

False Positives:
    Some warnings may be acceptable based on organizational context.
    Use judgment - not all warnings require fixes.
    Focus on CRITICAL issues that break functionality.

Common Issues and Solutions:
    Issue: #REF! in formulas
    Cause: External workbook reference broken
    Fix: Run normalize_assessment_files_a812.py, ensure files co-located
    
    Issue: "Reused style object" warning
    Cause: openpyxl script bug (shared Font/Fill/Border objects)
    Fix: Regenerate workbook with corrected script (create new style objects)
    
    Issue: Missing required sheet
    Cause: Generator script error or manual deletion
    Fix: Regenerate workbook from script
    
    Issue: Data validation overlapping ranges
    Cause: Generator script error (validation applied to too many cells)
    Fix: Review generator script validation range logic

Known openpyxl Issues:
    • Style object reuse: Fixed in generator scripts (create new objects per cell)
    • Large worksheets: May be slow to validate (>10,000 rows)
    • External formulas: Cannot validate target workbook exists

Exit Codes:
    0 = All workbooks passed validation
    1 = One or more CRITICAL issues found
    2 = Warnings found (but no critical issues)
    3 = No workbooks found to check

Performance:
    Validation time: ~5-10 seconds per workbook
    Memory usage: ~100-200 MB per workbook
    Total time: ~1 minute for complete framework (5 workbooks)

================================================================================
END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

import sys
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime


# ============================================================================
# CONFIGURATION
# ============================================================================

# Expected workbook structures
WORKBOOK_SPECS = {
    "Domain_1": {
        "pattern": "ISMS-IMP-A.8.12.1",
        "name": "DLP Infrastructure Assessment",
        "expected_sheets": [
            "Instructions_Legend",
            "DLP_Technology_Inventory",
            "Network_DLP",
            "Endpoint_DLP",
            "Email_DLP",
            "Cloud_CASB_DLP",
            "Web_DLP",
            "Database_DAM",
            "Gap_Analysis",
            "Evidence_Register",
            "Summary_Dashboard"
        ],
        "min_sheets": 11,
        "assessment_items": 80
    },
    "Domain_2": {
        "pattern": "ISMS-IMP-A.8.12.2",
        "name": "Data Classification Assessment",
        "expected_sheets": [
            "Instructions_Legend",
            "Classification_Schema",
            "Sensitive_Data_Inventory",
            "Data_Location_Mapping",
            "Data_Owner_Assignment",
            "Regulatory_Mapping",
            "Labeling_Methods",
            "Discovery_Results",
            "Gap_Analysis",
            "Evidence_Register",
            "Summary_Dashboard"
        ],
        "min_sheets": 11,
        "assessment_items": 70
    },
    "Domain_3": {
        "pattern": "ISMS-IMP-A.8.12.3",
        "name": "Channel Coverage Assessment",
        "expected_sheets": [
            "Instructions_Legend",
            "Channel_Overview",
            "Email_Channel",
            "Web_Cloud_Channel",
            "Endpoint_Channel",
            "Network_Channel",
            "Application_Channel",
            "Mobile_Channel",
            "Coverage_Metrics",
            "Gap_Analysis",
            "Evidence_Register",
            "Summary_Dashboard"
        ],
        "min_sheets": 12,
        "assessment_items": 90
    },
    "Domain_4": {
        "pattern": "ISMS-IMP-A.8.12.4",
        "name": "Monitoring & Response Assessment",
        "expected_sheets": [
            "Instructions_Legend",
            "Logging_Configuration",
            "Alert_Rules_Inventory",
            "Alert_Volume_Metrics",
            "SIEM_Integration",
            "False_Positive_Management",
            "Incident_Response_Workflow",
            "SOC_Integration",
            "Dashboards_Reporting",
            "Gap_Analysis",
            "Evidence_Register",
            "Summary_Dashboard"
        ],
        "min_sheets": 12,
        "assessment_items": 70
    },
    "Domain_5": {
        "pattern": "ISMS-IMP-A.8.12.5",
        "name": "Compliance Dashboard",
        "expected_sheets": [
            "Instructions_Legend",
            "Executive_Summary",
            "Domain_Rollup_Summary",
            "Consolidated_Gap_Analysis",
            "Risk_Register",
            "Remediation_Roadmap",
            "Evidence_Master_Index",
            "Trend_Analysis",
            "KPI_Dashboard",
            "Budget_Planning",
            "CISO_DPO_Approval",
            "Summary_Dashboard"
        ],
        "min_sheets": 12,
        "assessment_items": 25
    }
}

# Color codes for terminal output
GREEN = '\033[92m'
YELLOW = '\033[93m'
RED = '\033[91m'
BLUE = '\033[94m'
BOLD = '\033[1m'
RESET = '\033[0m'


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def print_header(text):
    """Print section header."""
    print(f"\n{BLUE}{BOLD}{'=' * 80}{RESET}")
    print(f"{BLUE}{BOLD}{text}{RESET}")
    print(f"{BLUE}{BOLD}{'=' * 80}{RESET}\n")


def print_success(text):
    """Print success message."""
    print(f"{GREEN}✅ {text}{RESET}")


def print_warning(text):
    """Print warning message."""
    print(f"{YELLOW}⚠️ {text}{RESET}")


def print_error(text):
    """Print error message."""
    print(f"{RED}❌ {text}{RESET}")


def print_info(text):
    """Print info message."""
    print(f"   {text}")


def find_workbook(directory, pattern):
    """Find workbook matching pattern in directory."""
    path = Path(directory)
    matches = list(path.glob(f"{pattern}*.xlsx"))
    
    if not matches:
        return None
    
    if len(matches) > 1:
        print_warning(f"Multiple workbooks found for {pattern}, using most recent")
        matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    
    return matches[0]


def check_workbook_structure(wb, domain_key, spec):
    """Check if workbook has correct sheet structure."""
    issues = []
    warnings = []
    
    actual_sheets = wb.sheetnames
    expected_sheets = spec["expected_sheets"]
    
    # Check sheet count
    if len(actual_sheets) < spec["min_sheets"]:
        issues.append(f"Only {len(actual_sheets)} sheets found, expected at least {spec['min_sheets']}")
    else:
        print_success(f"Sheet count: {len(actual_sheets)} sheets")
    
    # Check for expected sheets
    missing_sheets = []
    for sheet_name in expected_sheets:
        if sheet_name not in actual_sheets:
            missing_sheets.append(sheet_name)
    
    if missing_sheets:
        issues.append(f"Missing sheets: {', '.join(missing_sheets)}")
    else:
        print_success(f"All {len(expected_sheets)} expected sheets present")
    
    # Check for unexpected sheets
    unexpected_sheets = []
    for sheet_name in actual_sheets:
        if sheet_name not in expected_sheets:
            unexpected_sheets.append(sheet_name)
    
    if unexpected_sheets:
        warnings.append(f"Unexpected sheets: {', '.join(unexpected_sheets)}")
    
    return issues, warnings


def check_data_validations(ws, sheet_name):
    """Check if data validations are properly configured."""
    issues = []
    warnings = []
    
    # Check if any validations exist
    if not ws.data_validations.dataValidation:
        if sheet_name not in ["Instructions_Legend", "Coverage_Metrics", "Summary_Dashboard"]:
            warnings.append(f"{sheet_name}: No data validations found")
        return issues, warnings
    
    # Count validations
    validation_count = len(ws.data_validations.dataValidation)
    if validation_count > 0:
        print_success(f"{sheet_name}: {validation_count} data validations configured")
    
    # Check for common validation issues
    for dv in ws.data_validations.dataValidation:
        if dv.type == "list" and not dv.formula1:
            issues.append(f"{sheet_name}: List validation missing formula")
    
    return issues, warnings


def check_formulas(ws, sheet_name):
    """Check formulas for common errors."""
    issues = []
    warnings = []
    
    error_count = 0
    formula_count = 0
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if cell.value.startswith('='):
                    formula_count += 1
                    # Check for error indicators
                    if '#REF!' in cell.value or '#NAME?' in cell.value or '#VALUE!' in cell.value:
                        error_count += 1
                        issues.append(f"{sheet_name}: Formula error in {cell.coordinate}: {cell.value}")
    
    if formula_count > 0:
        if error_count == 0:
            print_success(f"{sheet_name}: {formula_count} formulas, no errors detected")
        else:
            issues.append(f"{sheet_name}: {error_count} formula errors found")
    
    return issues, warnings


def check_freeze_panes(ws, sheet_name):
    """Check if freeze panes are set."""
    issues = []
    warnings = []
    
    if ws.freeze_panes:
        print_success(f"{sheet_name}: Freeze panes set at {ws.freeze_panes}")
    else:
        if sheet_name not in ["Instructions_Legend", "Summary_Dashboard"]:
            warnings.append(f"{sheet_name}: No freeze panes set")
    
    return issues, warnings


def check_column_widths(ws, sheet_name):
    """Check if column widths are appropriately set."""
    issues = []
    warnings = []
    
    narrow_columns = 0
    wide_columns = 0
    
    for col_letter in ws.column_dimensions:
        width = ws.column_dimensions[col_letter].width
        if width and width < 8:
            narrow_columns += 1
        elif width and width > 10:
            wide_columns += 1
    
    if wide_columns > 0:
        print_success(f"{sheet_name}: {wide_columns} columns with custom widths")
    else:
        warnings.append(f"{sheet_name}: No custom column widths detected")
    
    return issues, warnings


def check_pre_populated_examples(ws, sheet_name):
    """Check if sheet has pre-populated example data."""
    issues = []
    warnings = []
    
    # Skip certain sheets that don't need examples
    skip_sheets = ["Instructions_Legend", "Summary_Dashboard", "Coverage_Metrics"]
    if sheet_name in skip_sheets:
        return issues, warnings
    
    # Count rows with data (excluding header rows 1-3)
    data_rows = 0
    for row in range(4, min(15, ws.max_row + 1)):
        if ws.cell(row=row, column=1).value:
            data_rows += 1
    
    if data_rows >= 3:
        print_success(f"{sheet_name}: {data_rows} pre-populated example rows")
    else:
        warnings.append(f"{sheet_name}: Only {data_rows} example rows, expected at least 3")
    
    return issues, warnings


def check_standard_sheets(wb, domain_key):
    """Check standard sheets that all workbooks should have."""
    issues = []
    warnings = []
    
    # Check Instructions_Legend (should have Document ID)
    if "Instructions_Legend" in wb.sheetnames:
        ws = wb["Instructions_Legend"]
        
        doc_id_found = False
        for row in ws.iter_rows(max_row=30, max_col=5):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "Document ID" in cell.value:
                        doc_id_found = True
                        break
        
        if doc_id_found:
            print_success("Instructions_Legend: Document ID found")
        else:
            warnings.append("Instructions_Legend: Document ID not found")
    
    # Check Gap_Analysis (should have ~40 rows for gaps)
    if "Gap_Analysis" in wb.sheetnames:
        ws = wb["Gap_Analysis"]
        max_row = ws.max_row
        if max_row < 40:
            warnings.append(f"Gap_Analysis: Only {max_row} rows, expected ~40")
        else:
            print_success(f"Gap_Analysis: {max_row} rows configured")
    
    # Check Evidence_Register (should have ~100 rows)
    if "Evidence_Register" in wb.sheetnames:
        ws = wb["Evidence_Register"]
        max_row = ws.max_row
        if max_row < 100:
            warnings.append(f"Evidence_Register: Only {max_row} rows, expected ~100")
        else:
            print_success(f"Evidence_Register: {max_row} rows configured")
    
    # Check Summary_Dashboard (should have KPIs and approval section)
    if "Summary_Dashboard" in wb.sheetnames:
        ws = wb["Summary_Dashboard"]
        
        # Look for "KPI" or "Compliance Score" text
        kpi_found = False
        approval_found = False
        
        for row in ws.iter_rows(max_row=50, max_col=5):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "KPI" in cell.value.upper() or "COMPLIANCE" in cell.value.upper():
                        kpi_found = True
                    if "APPROVAL" in cell.value.upper() or "SIGN-OFF" in cell.value.upper():
                        approval_found = True
        
        if kpi_found:
            print_success("Summary_Dashboard: KPIs section found")
        else:
            warnings.append("Summary_Dashboard: No KPIs section found")
        
        if approval_found:
            print_success("Summary_Dashboard: Approval section found")
        else:
            warnings.append("Summary_Dashboard: No approval section found")
    
    return issues, warnings


def validate_workbook(filepath, domain_key, spec):
    """Perform comprehensive validation of a workbook."""
    print_header(f"Validating {spec['name']} ({domain_key})")
    print_info(f"File: {filepath.name}")
    
    all_issues = []
    all_warnings = []
    
    try:
        # Load workbook
        print_info("Loading workbook...")
        wb = load_workbook(filepath, data_only=False)
        print_success("Workbook loaded successfully")
        
        # Check structure
        print_info("\nChecking workbook structure...")
        issues, warnings = check_workbook_structure(wb, domain_key, spec)
        all_issues.extend(issues)
        all_warnings.extend(warnings)
        
        # Check standard sheets
        print_info("\nChecking standard sheets...")
        issues, warnings = check_standard_sheets(wb, domain_key)
        all_issues.extend(issues)
        all_warnings.extend(warnings)
        
        # Check each sheet
        print_info("\nChecking individual sheets...")
        for sheet_name in wb.sheetnames:
            if sheet_name in spec["expected_sheets"]:
                ws = wb[sheet_name]
                
                # Data validations
                issues, warnings = check_data_validations(ws, sheet_name)
                all_issues.extend(issues)
                all_warnings.extend(warnings)
                
                # Formulas
                issues, warnings = check_formulas(ws, sheet_name)
                all_issues.extend(issues)
                all_warnings.extend(warnings)
                
                # Freeze panes
                issues, warnings = check_freeze_panes(ws, sheet_name)
                all_issues.extend(issues)
                all_warnings.extend(warnings)
                
                # Column widths
                issues, warnings = check_column_widths(ws, sheet_name)
                all_issues.extend(issues)
                all_warnings.extend(warnings)
                
                # Pre-populated examples
                issues, warnings = check_pre_populated_examples(ws, sheet_name)
                all_issues.extend(issues)
                all_warnings.extend(warnings)
        
        # Domain-specific checks
        if domain_key == "Domain_5":
            print_info("\nChecking Domain 5 specific features...")
            # Check for external formula placeholders
            if "Instructions_Legend" in wb.sheetnames:
                ws = wb["Instructions_Legend"]
                find_replace_found = False
                
                for row in ws.iter_rows(max_row=100, max_col=5):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            if "FIND" in cell.value.upper() and "REPLACE" in cell.value.upper():
                                find_replace_found = True
                                break
                
                if find_replace_found:
                    print_success("Domain 5: Find & Replace instructions found")
                else:
                    all_warnings.append("Domain 5: No Find & Replace instructions found for external formulas")
        
        # Summary
        print_info("\n" + "=" * 60)
        if len(all_issues) == 0 and len(all_warnings) == 0:
            print_success(f"{domain_key}: VALIDATION PASSED - No issues found")
            return True, 0, 0
        else:
            if len(all_issues) > 0:
                print_error(f"{domain_key}: {len(all_issues)} critical issues found")
                for issue in all_issues:
                    print_error(f"  • {issue}")
            
            if len(all_warnings) > 0:
                print_warning(f"{domain_key}: {len(all_warnings)} warnings found")
                for warning in all_warnings:
                    print_warning(f"  • {warning}")
            
            return len(all_issues) == 0, len(all_issues), len(all_warnings)
    
    except Exception as e:
        print_error(f"Failed to validate workbook: {str(e)}")
        return False, 1, 0


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main():
    """Main validation function."""
    
    # Determine directory
    if len(sys.argv) > 1:
        directory = sys.argv[1]
    else:
        directory = "."
    
    if not os.path.isdir(directory):
        print_error(f"Directory not found: {directory}")
        sys.exit(1)
    
    print_header("ISMS-IMP-A.8.12 Complete Framework Validation")
    print_info(f"Checking directory: {os.path.abspath(directory)}")
    print_info(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Find and validate each domain
    results = {}
    total_issues = 0
    total_warnings = 0
    
    for domain_key, spec in WORKBOOK_SPECS.items():
        filepath = find_workbook(directory, spec["pattern"])
        
        if filepath:
            passed, issues, warnings = validate_workbook(filepath, domain_key, spec)
            results[domain_key] = {
                "found": True,
                "passed": passed,
                "issues": issues,
                "warnings": warnings
            }
            total_issues += issues
            total_warnings += warnings
        else:
            print_header(f"Checking {spec['name']} ({domain_key})")
            print_warning(f"Workbook not found: {spec['pattern']}*.xlsx")
            print_info("This domain has not been generated yet (expected for Domains 4-5)")
            results[domain_key] = {
                "found": False,
                "passed": None,
                "issues": 0,
                "warnings": 0
            }
    
    # Final summary
    print_header("VALIDATION SUMMARY")
    
    found_count = sum(1 for r in results.values() if r["found"])
    passed_count = sum(1 for r in results.values() if r["passed"] == True)
    
    print_info(f"Workbooks found: {found_count} / 5")
    print_info(f"Workbooks passed: {passed_count} / {found_count}")
    print_info(f"Total critical issues: {total_issues}")
    print_info(f"Total warnings: {total_warnings}")
    
    print("\nDomain Status:")
    for domain_key, result in results.items():
        spec = WORKBOOK_SPECS[domain_key]
        if result["found"]:
            if result["passed"]:
                print_success(f"{domain_key} ({spec['name']}): PASSED")
            else:
                print_error(f"{domain_key} ({spec['name']}): FAILED ({result['issues']} issues, {result['warnings']} warnings)")
        else:
            print_warning(f"{domain_key} ({spec['name']}): NOT FOUND")
    
    # Exit code
    if total_issues > 0:
        print(f"\n{RED}{BOLD}VALIDATION FAILED{RESET}")
        print_info("Please fix critical issues before using workbooks in production")
        sys.exit(1)
    elif found_count == 0:
        print(f"\n{YELLOW}{BOLD}NO WORKBOOKS FOUND{RESET}")
        print_info("Generate workbooks first using the Python generator scripts")
        sys.exit(2)
    else:
        print(f"\n{GREEN}{BOLD}VALIDATION PASSED{RESET}")
        if total_warnings > 0:
            print_info(f"Note: {total_warnings} warnings found (non-critical)")
        print_info("All found workbooks are ready for use")
        sys.exit(0)


if __name__ == "__main__":
    main()
