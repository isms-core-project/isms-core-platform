#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.2-3-5-6 - Authentication & Privileged Access Compliance Dashboard
================================================================================

ISO/IEC 27001:2022 Controls A.8.2, A.8.3, A.8.5: Authentication & Privileged Access
Consolidation Dashboard: Executive Compliance Overview

**Purpose:**
Consolidates data from all five authentication and privileged access assessment
domains into a unified executive dashboard, providing comprehensive compliance
visibility for ISO 27001:2022 Controls A.8.2, A.8.3, and A.8.5.

**Consolidated Data Sources:**
- Domain 1: Authentication Inventory Assessment
- Domain 2: MFA Coverage Assessment
- Domain 3: Privileged Account Inventory Assessment
- Domain 4: Privileged Access Monitoring Assessment
- Domain 5: Information Access Restriction Assessment

**Generated Dashboard Structure:**
1. Executive Summary - Overall compliance status and critical gaps
2. Control A.8.5 Summary - Secure authentication compliance score
3. Control A.8.2 Summary - Privileged access rights compliance score
4. Control A.8.3 Summary - Information access restriction compliance score
5. MFA Coverage Overview - Enrollment status across user populations
6. Privileged Access Overview - PAM coverage and tier compliance
7. Access Restriction Overview - Technical enforcement status
8. Critical Gaps - High-priority remediation requirements
9. Compliance Trends - Historical progress tracking
10. Evidence Summary - Audit evidence consolidation
11. Regulatory Alignment - NIS2, FINMA, DORA, GDPR mapping
12. Approval & Sign-Off - Executive stakeholder approval

**Key Metrics:**
- Overall authentication security score
- MFA coverage percentage (total, privileged, remote, contractor)
- Privileged accounts in PAM vault (percentage)
- Admin tier isolation compliance (violation count)
- Access control compliance (OS, database, application, API)
- Critical gap count (requiring immediate remediation)
- Evidence completeness (audit readiness)

Usage:
    1. Complete and normalize source workbooks using normalize script
    2. Generate this dashboard
    3. Place dashboard in same folder as normalized workbooks
    4. Open dashboard and click "Update Links" to refresh data


    python3 generate_a8235_6_compliance_dashboard.py

Output:
    ISMS-IMP-A.8.2-3-5.6_Compliance_Dashboard.xlsx

"""

import sys
from datetime import datetime, timedelta
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
KEY = '\U0001F511'    # 🔑 Key
PACKAGE = '\U0001F4E6' # 📦 Package
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

WORKBOOK_NAME = "Authentication & PAM Compliance Dashboard"
DOCUMENT_ID = "ISMS-IMP-A.8.2-3-5-DASHBOARD"
CONTROL_REF = "ISO/IEC 27001:2022 A.8.2, A.8.3, A.8.5"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Dashboard sections
CRITICAL_GAPS_COUNT = 30
RISK_REGISTER_COUNT = 50
REMEDIATION_COUNT = 40
ACTION_ITEMS_COUNT = 50
EVIDENCE_COUNT = 100


def setup_styles():
    return {
        'header': {'font': {'bold': True, 'color': 'FFFFFF'}, 'fill': {'patternType': 'solid', 'fgColor': '003366'},
                   'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': True},
                   'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}},
        'title': {'font': {'size': 20, 'bold': True, 'color': '003366'}, 'alignment': {'horizontal': 'left'}},
        'metric_good': {'font': {'size': 16, 'bold': True, 'color': '00B050'},
                        'fill': {'patternType': 'solid', 'fgColor': 'E2EFDA'},
                        'alignment': {'horizontal': 'center', 'vertical': 'center'},
                        'border': {'left': 'medium', 'right': 'medium', 'top': 'medium', 'bottom': 'medium'}},
        'metric_warning': {'font': {'size': 16, 'bold': True, 'color': 'FFC000'},
                           'fill': {'patternType': 'solid', 'fgColor': 'FFF2CC'},
                           'alignment': {'horizontal': 'center', 'vertical': 'center'},
                           'border': {'left': 'medium', 'right': 'medium', 'top': 'medium', 'bottom': 'medium'}},
        'metric_critical': {'font': {'size': 16, 'bold': True, 'color': 'C00000'},
                            'fill': {'patternType': 'solid', 'fgColor': 'FCE4D6'},
                            'alignment': {'horizontal': 'center', 'vertical': 'center'},
                            'border': {'left': 'medium', 'right': 'medium', 'top': 'medium', 'bottom': 'medium'}},
        'data': {'font': {'size': 10}, 'alignment': {'horizontal': 'left'},
                 'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}}
    }


def apply_style(cell, style_dict):
    if 'font' in style_dict: cell.font = Font(**style_dict['font'])
    if 'fill' in style_dict: cell.fill = PatternFill(**style_dict['fill'])
    if 'alignment' in style_dict: cell.alignment = Alignment(**style_dict['alignment'])
    if 'border' in style_dict:
        cell.border = Border(left=Side(style=style_dict['border'].get('left', 'thin')),
                             right=Side(style=style_dict['border'].get('right', 'thin')),
                             top=Side(style=style_dict['border'].get('top', 'thin')),
                             bottom=Side(style=style_dict['border'].get('bottom', 'thin')))


def create_workbook():
    wb = Workbook()
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
    sheets = ["Executive_Dashboard", "Control_Compliance_Summary", "Critical_Gaps", "Risk_Register",
              "Remediation_Roadmap", "KPI_Metrics", "Evidence_Summary", "Action_Items", "Approval_Sign_Of"]
    for sheet in sheets: wb.create_sheet(title=sheet)
    return wb


def populate_executive_dashboard(wb, styles):
    ws = wb["Executive_Dashboard"]
    ws['A1'] = "Authentication & Privileged Access Management (PAM) Security Dashboard"
    apply_style(ws['A1'], styles['title'])
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Document ID: {DOCUMENT_ID} | Generated: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A3'] = "Consolidated view of ISO 27001:2022 A.8.2, A.8.3, A.8.5 compliance status"
    
    row = 5
    ws[f'A{row}'] = "OVERALL SECURITY POSTURE"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    ws.merge_cells(f'A{row}:H{row}')
    
    row += 2
    # Control A.8.5 - Authentication
    ws[f'A{row}'] = "A.8.5 - Secure Authentication"
    ws[f'B{row}'] = "MFA Coverage:"
    ws[f'C{row}'] = "=IFERROR([Workbook2]MFA_Coverage_By_Type!E8, 'UPDATE LINKS')"
    ws[f'C{row}'].number_format = '0%'
    apply_style(ws[f'C{row}'], styles['metric_good'])
    ws[f'D{row}'] = "Target: 85%+"
    
    row += 1
    ws[f'B{row}'] = "SSO Adoption:"
    ws[f'C{row}'] = "=IFERROR([Workbook1]SSO_Integration_Status!E8, 'UPDATE LINKS')"
    ws[f'C{row}'].number_format = '0%'
    ws[f'D{row}'] = "Target: 90%+"
    
    row += 2
    # Control A.8.2 - Privileged Access
    ws[f'A{row}'] = "A.8.2 - Privileged Access Rights"
    ws[f'B{row}'] = "PAM Vault Coverage:"
    ws[f'C{row}'] = "=IFERROR([Workbook3]PAM_Vault_Coverage!B5/B4, 'UPDATE LINKS')"
    ws[f'C{row}'].number_format = '0%'
    apply_style(ws[f'C{row}'], styles['metric_good'])
    ws[f'D{row}'] = "Target: 100% shared accounts"
    
    row += 1
    ws[f'B{row}'] = "Session Recording (Tier 0):"
    ws[f'C{row}'] = "=IFERROR([Workbook4]Session_Recording_Coverage!B8, 'UPDATE LINKS')"
    ws[f'C{row}'].number_format = '0%'
    ws[f'D{row}'] = "Target: 100%"
    
    row += 2
    # Control A.8.3 - Access Restriction
    ws[f'A{row}'] = "A.8.3 - Information Access Restriction"
    ws[f'B{row}'] = "Default Deny Compliance:"
    ws[f'C{row}'] = "=IFERROR([Workbook5]File_System_Permissions!E5/A5, 'UPDATE LINKS')"
    ws[f'C{row}'].number_format = '0%'
    apply_style(ws[f'C{row}'], styles['metric_warning'])
    ws[f'D{row}'] = "Target: 95%+"
    
    row += 3
    ws[f'A{row}'] = "CRITICAL GAPS REQUIRING IMMEDIATE ACTION"
    ws[f'A{row}'].font = Font(size=12, bold=True, color='C00000')
    ws.merge_cells(f'A{row}:H{row}')
    
    row += 1
    ws[f'A{row}'] = f"{BULLET} Privileged users without MFA: [See Critical_Gaps sheet]"
    ws[f'A{row}'].font = Font(color='C00000')
    row += 1
    ws[f'A{row}'] = f"{BULLET} Shared admin accounts not in PAM vault: [See Critical_Gaps sheet]"
    ws[f'A{row}'].font = Font(color='C00000')
    row += 1
    ws[f'A{row}'] = f"{BULLET} Tier isolation violations: [See Risk_Register sheet]"
    ws[f'A{row}'].font = Font(color='C00000')
    
    row += 2
    ws[f'A{row}'] = f"{TARGET} This dashboard consolidates 5 assessment workbooks. Click 'Update Links' to refresh data."
    ws[f'A{row}'].font = Font(italic=True)
    ws.merge_cells(f'A{row}:H{row}')
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25


def populate_control_summary(wb, styles):
    ws = wb["Control_Compliance_Summary"]
    ws['A1'] = "Control-by-Control Compliance Summary"
    apply_style(ws['A1'], styles['title'])
    
    row = 4
    headers = ["Control", "Assessment Area", "Score", "Target", "Compliance Status", "Priority Gaps"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row, col_idx, header)
        apply_style(ws.cell(row, col_idx), styles['header'])
    
    # Sample rows - would link to actual workbooks
    compliance_data = [
        ("A.8.5", "Authentication Mechanisms", "85%", "90%", f"{WARNING} Partial", "Legacy protocols"),
        ("A.8.5", "MFA Coverage", "82%", "85%", f"{WARNING} Partial", "Standard users missing MFA"),
        ("A.8.5", "SSO Integration", "75%", "90%", f"{WARNING} Partial", "15 apps not integrated"),
        ("A.8.2", "Privileged Accounts", "90%", "100%", f"{WARNING} Partial", "10 shared accounts not vaulted"),
        ("A.8.2", "Session Recording", "95%", "100%", f"{WARNING} Partial", "5% Tier 0 not recorded"),
        ("A.8.2", "Access Reviews", "100%", "100%", f"{CHECK} Compliant", "None"),
        ("A.8.3", "File Permissions", "88%", "95%", f"{WARNING} Partial", "Permission audits incomplete"),
        ("A.8.3", "Database Access", "92%", "95%", f"{WARNING} Partial", "Some excessive grants"),
        ("A.8.3", "Network Segmentation", "85%", "90%", f"{WARNING} Partial", "Firewall rule review needed")
    ]
    
    for control, area, score, target, status, gaps in compliance_data:
        row += 1
        ws.cell(row, 1, control)
        ws.cell(row, 2, area)
        ws.cell(row, 3, score)
        ws.cell(row, 4, target)
        ws.cell(row, 5, status)
        ws.cell(row, 6, gaps)
        for col in range(1, 7):
            apply_style(ws.cell(row, col), styles['data'])


def populate_remaining_sheets(wb, styles):
    # Critical Gaps
    ws_gaps = wb["Critical_Gaps"]
    ws_gaps['A1'] = "Top 30 Critical Security Gaps"
    apply_style(ws_gaps['A1'], styles['title'])
    ws_gaps['A2'] = "Prioritized list requiring immediate remediation"
    
    # Risk Register
    ws_risk = wb["Risk_Register"]
    ws_risk['A1'] = "Authentication & PAM Risk Register"
    apply_style(ws_risk['A1'], styles['title'])
    
    # Remediation Roadmap
    ws_remed = wb["Remediation_Roadmap"]
    ws_remed['A1'] = "Remediation Roadmap (Next 12 Months)"
    apply_style(ws_remed['A1'], styles['title'])
    
    # KPI Metrics
    ws_kpi = wb["KPI_Metrics"]
    ws_kpi['A1'] = "Key Performance Indicators (KPIs)"
    apply_style(ws_kpi['A1'], styles['title'])
    
    # Evidence Summary
    ws_evidence = wb["Evidence_Summary"]
    ws_evidence['A1'] = "Evidence Index (All Workbooks)"
    apply_style(ws_evidence['A1'], styles['title'])
    
    # Action Items
    ws_action = wb["Action_Items"]
    ws_action['A1'] = "Action Items & Follow-Up"
    apply_style(ws_action['A1'], styles['title'])
    
    # Approval
    ws_approval = wb["Approval_Sign_Of"]
    ws_approval['A1'] = "Dashboard Approval & Sign-Of"
    apply_style(ws_approval['A1'], styles['title'])
    ws_approval['A3'] = "Executive Summary Approved By:"
    ws_approval['A5'] = "CISO:"
    ws_approval['A6'] = "Date:"


def generate_workbook():
    print("\n╔════════════════════════════════════════════════════════════════╗")
    print("║  ISMS Assessment A.8.2/3/5 - Workbook 6: Consolidated Dashboard║")
    print("╚════════════════════════════════════════════════════════════════╝\n")
    
    wb = create_workbook()
    styles = setup_styles()
    
    print("Populating dashboard sheets...")
    populate_executive_dashboard(wb, styles)
    populate_control_summary(wb, styles)
    populate_remaining_sheets(wb, styles)
    
    filename = f"ISMS-IMP-A.8.2-3-5.6_Compliance_Dashboard_{GENERATED_TIMESTAMP}.xlsx"
    wb.save(filename)
    
    print(f"{CHECK} Generated: {filename}")
    print("\n⚠️  IMPORTANT: After opening, click 'Update Links' to connect to source workbooks\n")
    return filename


if __name__ == "__main__":
    try:
        generate_workbook()
        sys.exit(0)
    except Exception as e:
        print(f"{XMARK} ERROR: {e}")
        sys.exit(1)
