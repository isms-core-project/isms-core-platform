#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.2-3-5-3 - Privileged Account Inventory Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.2, A.8.3, A.8.5: Authentication & Privileged Access
Assessment Domain 3 of 6: Privileged Account Management

**Purpose:**
Enables systematic inventory and assessment of privileged accounts, admin
tiering compliance, and PAM implementation, supporting ISO 27001:2022 Control
A.8.2 privileged access requirements.

**Assessment Scope:**
- Privileged account inventory (named, shared, service, break-glass)
- Privileged user mapping and ownership
- Admin tier classification (Tier 0, Tier 1, Tier 2)
- Admin tier isolation compliance (cross-tier violation detection)
- PAM solution coverage (vaulted vs. non-vaulted accounts)
- Password rotation compliance
- MFA requirement compliance for privileged accounts
- Privileged account separation verification
- Dedicated admin workstation (PAW) deployment for Tier 0
- Gap analysis and remediation requirements
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Privileged access assessment guidance
2. Privileged Account Inventory - Complete account catalog with tier classification
3. Privileged User Mapping - User-to-account ownership and tier access
4. Admin Tier Compliance - Tier isolation violations and cross-tier access
5. PAM Coverage - Vaulted vs. non-vaulted account analysis
6. Password Rotation Status - Credential rotation compliance
7. MFA Compliance - Privileged account MFA enforcement
8. Tier 0 PAW Deployment - Dedicated workstation compliance
9. Gap Analysis - Non-compliant privileged accounts (prioritized)
10. Evidence Register - Audit evidence tracking
11. Approval & Sign-Off - Stakeholder review workflow

Usage:
    python3 generate_a8235_3_privileged_accounts.py

Requirements:
    - Python 3.7+
    - openpyxl library: pip install openpyxl

Output:
    ISMS-IMP-A.8.2-3-5.3_Privileged_Accounts_Assessment_YYYYMMDD.xlsx

"""

import sys
from datetime import datetime, timedelta
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================================
# SECTION 1: CONSTANTS AND CONFIGURATION
# ============================================================================



# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
KEY = '\U0001F511'    # 🔑 Key
PACKAGE = '\U0001F4E6' # 📦 Package
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

WORKBOOK_NAME = "Privileged Accounts Assessment"
DOCUMENT_ID = "ISMS-IMP-A.8.2-3-5.3"
VERSION = "1.0"
CONTROL_REF = "ISO/IEC 27001:2022 A.8.2 (Privileged Access Rights)"
GENERATED_DATE = datetime.now().strftime("%Y-%m-%d")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Workbook structure
ACCOUNT_ROW_COUNT = 150      # Pre-formatted rows for account inventory
USER_ROSTER_COUNT = 100      # Privileged user roster
TIER_MATRIX_COUNT = 30       # Tiering classification entries
PAM_COVERAGE_COUNT = 120     # PAM vault coverage tracking
TOKEN_DEPLOYMENT_COUNT = 40  # Hardware token tracking
ROTATION_STATUS_COUNT = 100  # Credential rotation tracking
REVIEW_RESULTS_COUNT = 80    # Quarterly review entries
TIER_VIOLATION_COUNT = 50    # Tier isolation violations
EVIDENCE_ROW_COUNT = 30      # Evidence register entries

# Admin Tiers
ADMIN_TIERS = [
    "Tier 0 (Domain/Enterprise/Critical)",
    "Tier 1 (Server/Application)",
    "Tier 2 (Workstation/Endpoint)",
    "N/A (Non-Privileged)"
]

# Account types
ACCOUNT_TYPES = [
    "Named Privileged (user.admin)",
    "Shared Privileged (root, Administrator, sa)",
    "Service Account (Privileged)",
    "Break-Glass / Emergency",
    "Cloud Global Admin",
    "Local Administrator",
    "Non-Privileged (Standard User)"
]

# Privileged roles
PRIVILEGED_ROLES = [
    "Domain Admin",
    "Enterprise Admin",
    "Schema Admin",
    "Azure Global Administrator",
    "AWS Root / Admin",
    "GCP Owner / Admin",
    "Database Administrator (DBA)",
    "Security Administrator",
    "Network Administrator",
    "Server Administrator",
    "Backup Administrator",
    "PKI Administrator",
    "PAM Administrator",
    "SIEM Administrator",
    "Workstation Local Admin",
    "Application Administrator",
    "Other (Specify)"
]

# MFA requirements for privileged
MFA_PRIVILEGED = [
    "🟢 Hardware Token (FIDO2) - Tier 0 Required",
    "🟢 Authenticator App (TOTP)",
    "🟡 Push Notification",
    f"{XMARK} Not Enrolled (Non-Compliant)"
]

# PAM status
PAM_STATUS = [
    f"{CHECK} Vaulted - Active",
    f"{CHECK} Vaulted - JIT Enabled",
    f"{WARNING} Vaulted - Rotation Issues",
    f"{XMARK} Not Vaulted (Gap)",
    "🔄 Onboarding In Progress",
    "➖ N/A (Named Account)"
]

# Credential rotation frequency
ROTATION_FREQUENCY = [
    "After Each Use (Check-in)",
    "Daily",
    "Weekly",
    "Monthly",
    "Quarterly",
    "On Demand Only",
    "Not Applicable"
]

# Compliance status
COMPLIANCE_STATUS = [
    f"{CHECK} Compliant",
    f"{WARNING} Partial Compliance",
    f"{XMARK} Non-Compliant",
    "🔄 In Progress",
    "📋 Under Review",
    "❓ Unknown",
    "➖ N/A"
]

# Priority levels
PRIORITY_LEVELS = [
    "🔴 Critical (Tier 0)",
    "🟠 High (Tier 1 Production)",
    "🟡 Medium (Tier 1 Non-Prod)",
    "🟢 Low (Tier 2)",
    "⚪ Informational"
]


# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================

def setup_styles():
    """Define all cell styles used throughout the workbook."""
    return {
        'header': {
            'font': {'name': 'Calibri', 'size': 11, 'bold': True, 'color': 'FFFFFF'},
            'fill': {'patternType': 'solid', 'fgColor': '003366'},
            'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'subheader': {
            'font': {'name': 'Calibri', 'size': 11, 'bold': True, 'color': '000000'},
            'fill': {'patternType': 'solid', 'fgColor': 'D9E1F2'},
            'alignment': {'horizontal': 'left', 'vertical': 'center', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'data': {
            'font': {'name': 'Calibri', 'size': 10, 'color': '000000'},
            'fill': {'patternType': 'solid', 'fgColor': 'FFFFFF'},
            'alignment': {'horizontal': 'left', 'vertical': 'center', 'wrap_text': False},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'title': {
            'font': {'name': 'Calibri', 'size': 16, 'bold': True, 'color': '003366'},
            'alignment': {'horizontal': 'left', 'vertical': 'center'}
        },
        'tier0': {
            'font': {'name': 'Calibri', 'size': 10, 'bold': True, 'color': '000000'},
            'fill': {'patternType': 'solid', 'fgColor': 'FCE4D6'},
            'alignment': {'horizontal': 'left', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'tier1': {
            'font': {'name': 'Calibri', 'size': 10, 'color': '000000'},
            'fill': {'patternType': 'solid', 'fgColor': 'FFF2CC'},
            'alignment': {'horizontal': 'left', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'tier2': {
            'font': {'name': 'Calibri', 'size': 10, 'color': '000000'},
            'fill': {'patternType': 'solid', 'fgColor': 'E2EFDA'},
            'alignment': {'horizontal': 'left', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        }
    }


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell with NEW objects."""
    if 'font' in style_dict:
        cell.font = Font(**style_dict['font'])
    if 'fill' in style_dict:
        cell.fill = PatternFill(**style_dict['fill'])
    if 'alignment' in style_dict:
        cell.alignment = Alignment(**style_dict['alignment'])
    if 'border' in style_dict:
        cell.border = Border(
            left=Side(style=style_dict['border'].get('left', 'thin')),
            right=Side(style=style_dict['border'].get('right', 'thin')),
            top=Side(style=style_dict['border'].get('top', 'thin')),
            bottom=Side(style=style_dict['border'].get('bottom', 'thin'))
        )


# ============================================================================
# SECTION 3: WORKBOOK CREATION
# ============================================================================

def create_workbook():
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create sheets matching specification
    sheets = [
        "Instructions_Legend",
        "Privileged_Account_Inventory",
        "Admin_Tiering_Matrix",
        "Privileged_User_Roster",
        "PAM_Vault_Coverage",
        "MFA_Hardware_Tokens",
        "Credential_Rotation_Status",
        "Access_Review_Results",
        "Tier_Isolation_Compliance",
        "Evidence_Register",
        "Approval_Sign_Of"
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    
    return wb


# ============================================================================
# SECTION 4: SHEET 1 - INSTRUCTIONS & LEGEND
# ============================================================================

def populate_instructions(wb, styles):
    """Populate Instructions & Legend sheet with admin tiering explanation."""
    ws = wb["Instructions_Legend"]
    
    # Title
    ws['A1'] = "Privileged Accounts Inventory & Admin Tiering Assessment"
    apply_style(ws['A1'], styles['title'])
    
    # Document metadata (standardized rows 3-6)
    ws['A3'] = 'Document ID:'
    ws['A3'].font = Font(bold=True)
    ws['B3'] = DOCUMENT_ID
    
    ws['A4'] = 'Assessment:'
    ws['A4'].font = Font(bold=True)
    ws['B4'] = 'Privileged Accounts Inventory'
    
    ws['A5'] = 'Version:'
    ws['A5'].font = Font(bold=True)
    ws['B5'] = VERSION
    
    ws['A6'] = 'Generated:'
    ws['A6'].font = Font(bold=True)
    ws['B6'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    ws.column_dimensions['B'].width = 40
    
    # Critical Security Notice
    row = 8
    ws[f'A{row}'] = f"{WARNING} PRIVILEGED ACCESS = HIGHEST SECURITY RISK"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='C00000')
    ws[f'A{row}'].fill = PatternFill(patternType='solid', fgColor='FCE4D6')
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    ws[f'A{row}'] = "80% of data breaches involve privileged access abuse. Privileged accounts have access to critical infrastructure and sensitive data. Compromise of privileged credentials = complete organizational compromise. Admin Tiering Model (Tier 0/1/2) prevents lateral movement and limits blast radius."
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(f'A{row}:H{row}')
    ws.row_dimensions[row].height = 60
    
    row += 2
    ws[f'A{row}'] = "ADMIN TIERING MODEL (TIER 0 / 1 / 2)"
    apply_style(ws[f'A{row}'], styles['subheader'])
    row += 1
    ws[f'A{row}'] = "The Admin Tiering Model prevents credential theft and lateral movement by isolating privileged accounts into three tiers with strict separation. Tier 0 accounts NEVER log into Tier 1 or Tier 2 systems."
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(f'A{row}:H{row}')
    ws.row_dimensions[row].height = 35
    
    row += 2
    ws[f'A{row}'] = "TIER 0: DOMAIN / ENTERPRISE / CRITICAL INFRASTRUCTURE"
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='C00000')
    ws[f'A{row}'].fill = PatternFill(patternType='solid', fgColor='FCE4D6')
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    tier0_info = [
        f"{BULLET} Systems: Domain Controllers, Azure AD tenant, AWS Organization root, PKI CA, SIEM, PAM vault, Backup infrastructure",
        f"{BULLET} Accounts: Domain Admins, Enterprise Admins, Global Administrator, AWS root, Security admins, PAM admins",
        f"{BULLET} Access Level: Can access EVERYTHING - highest privilege in organization",
        f"{BULLET} Security Requirements:",
        "  - Dedicated PAWs (Privileged Access Workstations) REQUIRED",
        "  - Hardware MFA tokens (FIDO2) MANDATORY",
        "  - Session recording 100%",
        "  - JIT (Just-in-Time) access preferred",
        "  - Network isolation (PAWs can ONLY access Tier 0 systems)",
        "  - NO internet access, NO email, NO web browsing from PAWs",
        "  - Quarterly access reviews MANDATORY",
        f"{BULLET} Isolation Rule: Tier 0 accounts NEVER log into Tier 1 or Tier 2 systems (prevents credential theft)"
    ]
    
    for info in tier0_info:
        ws[f'A{row}'] = info
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = 22
        row += 1
    
    row += 1
    ws[f'A{row}'] = "TIER 1: SERVER / APPLICATION INFRASTRUCTURE"
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFC000')
    ws[f'A{row}'].fill = PatternFill(patternType='solid', fgColor='FFF2CC')
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    tier1_info = [
        f"{BULLET} Systems: Production servers, databases, ERP/CRM apps, virtualization, cloud infrastructure (non-global)",
        f"{BULLET} Accounts: Server admins, DBAs, application admins, VMware admins, Azure subscription admins",
        f"{BULLET} Access Level: Can manage servers and applications, CANNOT access Tier 0 or Tier 2",
        f"{BULLET} Security Requirements:",
        "  - Separate admin accounts from standard user accounts",
        "  - MFA MANDATORY (authenticator app minimum)",
        "  - Session recording for production systems",
        "  - Quarterly access reviews",
        "  - PAM vaulting for shared accounts",
        f"{BULLET} Isolation Rule: Tier 1 accounts CANNOT log into Tier 0 or Tier 2 systems"
    ]
    
    for info in tier1_info:
        ws[f'A{row}'] = info
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = 22
        row += 1
    
    row += 1
    ws[f'A{row}'] = "TIER 2: WORKSTATION / ENDPOINT"
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='00B050')
    ws[f'A{row}'].fill = PatternFill(patternType='solid', fgColor='E2EFDA')
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    tier2_info = [
        f"{BULLET} Systems: End-user workstations, laptops, mobile devices",
        f"{BULLET} Accounts: Desktop support, help desk with local admin, MDM administrators",
        f"{BULLET} Access Level: Can manage user endpoints ONLY, CANNOT access Tier 0 or Tier 1",
        f"{BULLET} Security Requirements:",
        "  - Local admin only (no domain admin privileges)",
        "  - MFA MANDATORY",
        "  - Standard logging",
        f"{BULLET} Isolation Rule: Tier 2 accounts CANNOT access Tier 0 or Tier 1 systems"
    ]
    
    for info in tier2_info:
        ws[f'A{row}'] = info
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = 22
        row += 1
    
    row += 2
    ws[f'A{row}'] = "TIERING BENEFITS"
    apply_style(ws[f'A{row}'], styles['subheader'])
    row += 1
    
    benefits = [
        f"{CHECK} Prevents Lateral Movement: Attacker compromises Tier 2 workstation → CANNOT reach Tier 0 domain controllers",
        f"{CHECK} Limits Blast Radius: Credential compromise contained to single tier",
        f"{CHECK} Enforces Least Privilege: Admins only have privileges needed for their tier",
        f"{CHECK} Architectural Security: Security enforced at network/GPO level, not just policy"
    ]
    
    for benefit in benefits:
        ws[f'A{row}'] = benefit
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = 22
        row += 1
    
    row += 2
    ws[f'A{row}'] = "PRIVILEGED ACCOUNT REQUIREMENTS"
    apply_style(ws[f'A{row}'], styles['subheader'])
    row += 1
    
    requirements = [
        f"{BULLET} Named Privileged Accounts: user.tier0, user.tier1, user.tier2 (separate from standard user.name account)",
        f"{BULLET} Shared Accounts (root, Administrator, sa): 100% MUST be in PAM vault",
        f"{BULLET} MFA: 100% of privileged users MUST have MFA (hardware tokens for Tier 0)",
        f"{BULLET} PAM Vaulting: All shared privileged accounts vaulted with automated password rotation",
        f"{BULLET} Session Recording: 100% Tier 0, 90%+ Tier 1 production",
        f"{BULLET} Access Reviews: Quarterly reviews MANDATORY (who has privileged access, still justified?)",
        f"{BULLET} Break-Glass Accounts: Sealed credentials, dual custody, usage triggers immediate review"
    ]
    
    for req in requirements:
        ws[f'A{row}'] = req
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = 22
        row += 1
    
    row += 2
    ws[f'A{row}'] = f"{TARGET} REMEMBER: Privileged access is the 'keys to the kingdom'. Get this wrong and no other security controls matter."
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, italic=True, color='C00000')
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='center')
    ws.merge_cells(f'A{row}:H{row}')
    ws.row_dimensions[row].height = 30
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 20


# ============================================================================
# SECTION 5: SHEET 2 - PRIVILEGED ACCOUNT INVENTORY
# ============================================================================

def populate_account_inventory(wb, styles):
    """Populate Privileged Account Inventory sheet."""
    ws = wb["Privileged_Account_Inventory"]
    
    # Title
    ws['A1'] = "Privileged Account Inventory"
    apply_style(ws['A1'], styles['title'])
    ws.merge_cells('A1:N1')
    
    ws['A2'] = "Complete inventory of all privileged accounts with tier classification"
    ws.merge_cells('A2:N2')
    
    # Headers
    headers = [
        "Account Name",
        "Account Type",
        "Admin Tier",
        "Privileged Role",
        "System / Platform",
        "Owner / Responsible",
        "MFA Status",
        "MFA Method",
        "PAM Vaulted",
        "Last Password Change",
        "Password Rotation Freq",
        "Last Access Review",
        "Compliance Status",
        "Notes"
    ]
    
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    
    # Pre-format data rows
    for row_num in range(5, 5 + ACCOUNT_ROW_COUNT):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    
    # Add dropdowns
    # Column B: Account Type
    dv_account_type = DataValidation(type="list", formula1=f'"{",".join(ACCOUNT_TYPES)}"', allow_blank=True)
    ws.add_data_validation(dv_account_type)
    dv_account_type.add(f'B5:B{4 + ACCOUNT_ROW_COUNT}')
    
    # Column C: Admin Tier
    dv_tier = DataValidation(type="list", formula1=f'"{",".join(ADMIN_TIERS)}"', allow_blank=True)
    ws.add_data_validation(dv_tier)
    dv_tier.add(f'C5:C{4 + ACCOUNT_ROW_COUNT}')
    
    # Column D: Privileged Role
    dv_role = DataValidation(type="list", formula1=f'"{",".join(PRIVILEGED_ROLES)}"', allow_blank=True)
    ws.add_data_validation(dv_role)
    dv_role.add(f'D5:D{4 + ACCOUNT_ROW_COUNT}')
    
    # Column G: MFA Status
    dv_mfa_status = DataValidation(type="list", formula1='f"{CHECK} Enrolled,❌ Not Enrolled"', allow_blank=True)
    ws.add_data_validation(dv_mfa_status)
    dv_mfa_status.add(f'G5:G{4 + ACCOUNT_ROW_COUNT}')
    
    # Column H: MFA Method
    dv_mfa_method = DataValidation(type="list", formula1=f'"{",".join(MFA_PRIVILEGED)}"', allow_blank=True)
    ws.add_data_validation(dv_mfa_method)
    dv_mfa_method.add(f'H5:H{4 + ACCOUNT_ROW_COUNT}')
    
    # Column I: PAM Vaulted
    dv_pam = DataValidation(type="list", formula1=f'"{",".join(PAM_STATUS)}"', allow_blank=True)
    ws.add_data_validation(dv_pam)
    dv_pam.add(f'I5:I{4 + ACCOUNT_ROW_COUNT}')
    
    # Column K: Rotation Frequency
    dv_rotation = DataValidation(type="list", formula1=f'"{",".join(ROTATION_FREQUENCY)}"', allow_blank=True)
    ws.add_data_validation(dv_rotation)
    dv_rotation.add(f'K5:K{4 + ACCOUNT_ROW_COUNT}')
    
    # Column M: Compliance Status
    dv_compliance = DataValidation(type="list", formula1=f'"{",".join(COMPLIANCE_STATUS)}"', allow_blank=True)
    ws.add_data_validation(dv_compliance)
    dv_compliance.add(f'M5:M{4 + ACCOUNT_ROW_COUNT}')
    
    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 32
    ws.column_dimensions['I'].width = 22
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 22
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 18
    ws.column_dimensions['N'].width = 35
    
    # Freeze panes
    ws.freeze_panes = 'A5'


# ============================================================================
# SECTION 6: SHEET 3 - ADMIN TIERING MATRIX
# ============================================================================

def populate_tiering_matrix(wb, styles):
    """Populate Admin Tiering Matrix sheet."""
    ws = wb["Admin_Tiering_Matrix"]
    
    # Title
    ws['A1'] = "Admin Tiering Classification Matrix"
    apply_style(ws['A1'], styles['title'])
    ws.merge_cells('A1:H1')
    
    ws['A2'] = "Classify systems and accounts into Tier 0/1/2 for isolation enforcement"
    ws.merge_cells('A2:H2')
    
    # Tier 0 Systems Section
    row = 4
    ws[f'A{row}'] = "TIER 0 SYSTEMS (Domain/Enterprise/Critical)"
    apply_style(ws[f'A{row}'], styles['tier0'])
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = "System / Infrastructure"
    ws[f'B{row}'] = "Tier Classification"
    ws[f'C{row}'] = "Criticality"
    ws[f'D{row}'] = "Security Requirements"
    ws[f'E{row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['header'])
    
    tier0_systems = [
        ("Domain Controllers", "Tier 0", "Critical", "PAWs, Hardware MFA, Session Recording", ""),
        ("Azure AD Tenant", "Tier 0", "Critical", "Global Admin requires PAW", ""),
        ("AWS Organization Root", "Tier 0", "Critical", "Root user MFA, breakglass only", ""),
        ("PKI Certificate Authority", "Tier 0", "Critical", "CA admins Tier 0", ""),
        ("SIEM Infrastructure", "Tier 0", "Critical", "Security admins Tier 0", ""),
        ("PAM Vault", "Tier 0", "Critical", "PAM admins Tier 0", ""),
        ("Backup Infrastructure", "Tier 0", "Critical", "Backup admins Tier 0", "")
    ]
    
    for system, tier, criticality, requirements, notes in tier0_systems:
        row += 1
        ws[f'A{row}'] = system
        ws[f'B{row}'] = tier
        ws[f'C{row}'] = criticality
        ws[f'D{row}'] = requirements
        ws[f'E{row}'] = notes
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles['tier0'])
    
    # Tier 1 Systems Section
    row += 2
    ws[f'A{row}'] = "TIER 1 SYSTEMS (Server/Application)"
    apply_style(ws[f'A{row}'], styles['tier1'])
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = "System / Infrastructure"
    ws[f'B{row}'] = "Tier Classification"
    ws[f'C{row}'] = "Criticality"
    ws[f'D{row}'] = "Security Requirements"
    ws[f'E{row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['header'])
    
    tier1_systems = [
        ("Production Servers", "Tier 1", "High", "Separate admin accounts, MFA, Session recording", ""),
        ("Production Databases", "Tier 1", "High", "DBAs Tier 1, PAM vaulting", ""),
        ("ERP / CRM Systems", "Tier 1", "High", "App admins Tier 1", ""),
        ("Cloud VMs (Non-Global)", "Tier 1", "Medium", "Subscription admins Tier 1", ""),
        ("Test/Dev Servers", "Tier 1", "Medium", "Same admin accounts as prod", "")
    ]
    
    for system, tier, criticality, requirements, notes in tier1_systems:
        row += 1
        ws[f'A{row}'] = system
        ws[f'B{row}'] = tier
        ws[f'C{row}'] = criticality
        ws[f'D{row}'] = requirements
        ws[f'E{row}'] = notes
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles['tier1'])
    
    # Tier 2 Systems Section
    row += 2
    ws[f'A{row}'] = "TIER 2 SYSTEMS (Workstation/Endpoint)"
    apply_style(ws[f'A{row}'], styles['tier2'])
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = "System / Infrastructure"
    ws[f'B{row}'] = "Tier Classification"
    ws[f'C{row}'] = "Criticality"
    ws[f'D{row}'] = "Security Requirements"
    ws[f'E{row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['header'])
    
    tier2_systems = [
        ("User Workstations", "Tier 2", "Low", "Local admin only, MFA", ""),
        ("User Laptops", "Tier 2", "Low", "Help desk local admin", ""),
        ("Mobile Devices", "Tier 2", "Low", "MDM admins", "")
    ]
    
    for system, tier, criticality, requirements, notes in tier2_systems:
        row += 1
        ws[f'A{row}'] = system
        ws[f'B{row}'] = tier
        ws[f'C{row}'] = criticality
        ws[f'D{row}'] = requirements
        ws[f'E{row}'] = notes
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles['tier2'])
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 30


# ============================================================================
# SECTION 7: REMAINING SHEETS (USER ROSTER, PAM, MFA, ROTATION, REVIEWS, ETC.)
# ============================================================================

def populate_remaining_sheets(wb, styles):
    """Populate remaining sheets with standard structure."""
    
    # Sheet 4: Privileged User Roster
    ws_roster = wb["Privileged_User_Roster"]
    ws_roster['A1'] = "Privileged User Roster"
    apply_style(ws_roster['A1'], styles['title'])
    ws_roster['A2'] = "Mapping of users to privileged accounts and tier access"
    
    roster_headers = ["User Name", "User ID", "Privileged Accounts Owned", "Tier 0 Access", "Tier 1 Access", "Tier 2 Access", "Business Justification", "Last Review Date"]
    row = 4
    for col_idx, header in enumerate(roster_headers, start=1):
        cell = ws_roster.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    
    for row_num in range(5, 5 + USER_ROSTER_COUNT):
        for col_idx in range(1, len(roster_headers) + 1):
            cell = ws_roster.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    
    # Sheet 5: PAM Vault Coverage
    ws_pam = wb["PAM_Vault_Coverage"]
    ws_pam['A1'] = "PAM Vault Coverage Tracking"
    apply_style(ws_pam['A1'], styles['title'])
    ws_pam['A2'] = "Track which privileged accounts are vaulted in PAM solution (Target: 100% shared accounts)"
    
    pam_headers = ["Account Name", "Account Type", "PAM Status", "Vault Date", "JIT Enabled", "Session Recording", "Auto Rotation", "Compliance"]
    row = 4
    for col_idx, header in enumerate(pam_headers, start=1):
        cell = ws_pam.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    
    for row_num in range(5, 5 + PAM_COVERAGE_COUNT):
        for col_idx in range(1, len(pam_headers) + 1):
            cell = ws_pam.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    
    # Sheet 6: MFA Hardware Tokens
    ws_tokens = wb["MFA_Hardware_Tokens"]
    ws_tokens['A1'] = "Hardware Token Deployment (Tier 0 Requirement)"
    apply_style(ws_tokens['A1'], styles['title'])
    ws_tokens['A2'] = "Track hardware token (FIDO2) deployment for Tier 0 admins"
    
    token_headers = ["User Name", "Tier 0 Account", "Primary Token Serial", "Backup Token Serial", "Enrollment Date", "Last Tested", "Status"]
    row = 4
    for col_idx, header in enumerate(token_headers, start=1):
        cell = ws_tokens.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    
    for row_num in range(5, 5 + TOKEN_DEPLOYMENT_COUNT):
        for col_idx in range(1, len(token_headers) + 1):
            cell = ws_tokens.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    
    # Sheet 7: Credential Rotation Status
    ws_rotation = wb["Credential_Rotation_Status"]
    ws_rotation['A1'] = "Credential Rotation Compliance"
    apply_style(ws_rotation['A1'], styles['title'])
    ws_rotation['A2'] = "Track password rotation for privileged accounts"
    
    rotation_headers = ["Account Name", "Last Rotation", "Rotation Frequency", "Next Due", "Days Until Due", "Rotation Method", "Compliance"]
    row = 4
    for col_idx, header in enumerate(rotation_headers, start=1):
        cell = ws_rotation.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    
    for row_num in range(5, 5 + ROTATION_STATUS_COUNT):
        for col_idx in range(1, len(rotation_headers) + 1):
            cell = ws_rotation.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
        
        # Formula for Days Until Due (Column E)
        e_cell = ws_rotation.cell(row=row_num, column=5)
        e_cell.value = f'=IF(D{row_num}<>"",D{row_num}-TODAY(),"")'
    
    # Sheet 8: Access Review Results
    ws_review = wb["Access_Review_Results"]
    ws_review['A1'] = "Quarterly Privileged Access Reviews"
    apply_style(ws_review['A1'], styles['title'])
    ws_review['A2'] = "Track quarterly reviews of privileged access (MANDATORY)"
    
    review_headers = ["Review Period", "User", "Privileged Account", "Access Confirmed", "Access Removed", "Access Justified (Exception)", "Reviewer", "Review Date"]
    row = 4
    for col_idx, header in enumerate(review_headers, start=1):
        cell = ws_review.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    
    for row_num in range(5, 5 + REVIEW_RESULTS_COUNT):
        for col_idx in range(1, len(review_headers) + 1):
            cell = ws_review.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    
    # Sheet 9: Tier Isolation Compliance
    ws_tier = wb["Tier_Isolation_Compliance"]
    ws_tier['A1'] = "Tier Isolation Violation Tracking"
    apply_style(ws_tier['A1'], styles['title'])
    ws_tier['A2'] = "Monitor and investigate cross-tier login attempts (should be ZERO)"
    
    tier_headers = ["Date", "Account", "Account Tier", "System Accessed", "System Tier", "Violation Type", "Investigation Status", "Remediation"]
    row = 4
    for col_idx, header in enumerate(tier_headers, start=1):
        cell = ws_tier.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    
    for row_num in range(5, 5 + TIER_VIOLATION_COUNT):
        for col_idx in range(1, len(tier_headers) + 1):
            cell = ws_tier.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    
    # Sheet 10: Evidence Register
    ws_evidence = wb["Evidence_Register"]
    ws_evidence['A1'] = "Privileged Access Evidence Register"
    apply_style(ws_evidence['A1'], styles['title'])
    
    evidence_headers = ["Evidence ID", "Evidence Type", "Description", "Source", "Date", "Collected By", "Verification"]
    row = 4
    for col_idx, header in enumerate(evidence_headers, start=1):
        cell = ws_evidence.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    
    for row_num in range(5, 5 + EVIDENCE_ROW_COUNT):
        for col_idx in range(1, len(evidence_headers) + 1):
            cell = ws_evidence.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    
    # Sheet 11: Approval Sign-Off
    ws_approval = wb["Approval_Sign_Of"]
    ws_approval['A1'] = "Privileged Access Assessment Approval"
    apply_style(ws_approval['A1'], styles['title'])
    
    ws_approval['A3'] = "Assessment Completion"
    apply_style(ws_approval['A3'], styles['subheader'])
    ws_approval['A4'] = "Assessment Date:"
    ws_approval['A5'] = "Assessed By:"
    ws_approval['A6'] = "Privileged Account Count:"
    
    ws_approval['A8'] = "Approval Workflow"
    apply_style(ws_approval['A8'], styles['subheader'])
    
    approval_rows = [
        ("Level 1", "Security Architect", "", "", ""),
        ("Level 2", "IAM Lead", "", "", ""),
        ("Level 3", "CISO", "", "", "")
    ]
    
    row = 9
    ws_approval[f'A{row}'] = "Level"
    ws_approval[f'B{row}'] = "Role"
    ws_approval[f'C{row}'] = "Name"
    ws_approval[f'D{row}'] = "Date"
    ws_approval[f'E{row}'] = "Signature"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws_approval[f'{col}{row}'], styles['header'])
    
    for level, role, name, date, sig in approval_rows:
        row += 1
        ws_approval[f'A{row}'] = level
        ws_approval[f'B{row}'] = role
        ws_approval[f'C{row}'] = name
        ws_approval[f'D{row}'] = date
        ws_approval[f'E{row}'] = sig


# ============================================================================
# SECTION 8: MAIN GENERATION FUNCTION
# ============================================================================

def generate_workbook():
    """Main function to generate complete workbook."""
    print()
    print("╔════════════════════════════════════════════════════════════════╗")
    print("║  ISMS Assessment A.8.2/3/5 - Authentication & PAM Framework    ║")
    print("║  Workbook 3: Privileged Accounts & Admin Tiering Assessment    ║")
    print("║                                                                ║")
    print("║  'Privileged access = keys to the kingdom'                     ║")
    print("╚════════════════════════════════════════════════════════════════╝")
    print()
    
    print("Creating workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    
    print("Populating Instructions & Legend (Admin Tiering Model)...")
    populate_instructions(wb, styles)
    
    print("Populating Privileged Account Inventory (150 accounts)...")
    populate_account_inventory(wb, styles)
    
    print("Populating Admin Tiering Classification Matrix...")
    populate_tiering_matrix(wb, styles)
    
    print("Populating remaining sheets...")
    populate_remaining_sheets(wb, styles)
    
    # Save workbook
    filename = f"ISMS-IMP-A.8.2-3-5.3_Privileged_Accounts_{GENERATED_TIMESTAMP}.xlsx"
    wb.save(filename)
    
    print()
    print("=" * 70)
    print(f"{CHECK} Workbook generated successfully: {filename}")
    print("=" * 70)
    print()
    print("Next Steps:")
    print("  1. Discover all privileged accounts (AD groups, cloud admins, DBAs)")
    print("  2. Classify accounts into Tier 0/1/2 in Admin_Tiering_Matrix")
    print("  3. Complete Privileged_Account_Inventory with tier assignments")
    print("  4. Verify PAM_Vault_Coverage (target: 100% shared accounts)")
    print("  5. Track MFA_Hardware_Tokens for Tier 0 admins")
    print("  6. Monitor Credential_Rotation_Status")
    print("  7. Conduct quarterly Access_Review_Results")
    print("  8. Investigate any Tier_Isolation_Compliance violations")
    print("  9. Collect evidence in Evidence_Register")
    print(" 10. Obtain approvals in Approval_Sign_Of")
    print()
    print("Critical Requirements:")
    print("  • Tier 0: Hardware MFA tokens MANDATORY, PAWs required")
    print("  • All Privileged: 100% MFA coverage (no exceptions)")
    print("  • Shared Accounts: 100% in PAM vault")
    print("  • Tier Isolation: ZERO cross-tier violations")
    print("  • Access Reviews: Quarterly MANDATORY")
    print()
    
    return filename


def validate_workbook(filename):
    """Validate generated workbook."""
    print("Running validation...")
    print("-" * 70)
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        
        expected_sheets = [
            "Instructions_Legend",
            "Privileged_Account_Inventory",
            "Admin_Tiering_Matrix",
            "Privileged_User_Roster",
            "PAM_Vault_Coverage",
            "MFA_Hardware_Tokens",
            "Credential_Rotation_Status",
            "Access_Review_Results",
            "Tier_Isolation_Compliance",
            "Evidence_Register",
            "Approval_Sign_Of"
        ]
        
        # Check sheet count
        if len(wb.sheetnames) != 11:
            print(f"  ❌ Expected 11 sheets, found {len(wb.sheetnames)}")
            return False
        print(f"  ✅ Sheet count: {len(wb.sheetnames)}")
        
        # Check sheet names
        for sheet in expected_sheets:
            if sheet in wb.sheetnames:
                print(f"  ✅ Found: {sheet}")
            else:
                print(f"  ❌ Missing: {sheet}")
                return False
        
        print()
        print("Validation Result: ✅ PASSED")
        print()
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"  ❌ Validation error: {str(e)}")
        return False


# ============================================================================
# SECTION 9: ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    """Entry point for script execution."""
    import sys
    
    try:
        # Generate the workbook
        output_file = generate_workbook()
        
        # Validate the output
        if validate_workbook(output_file):
            print(f"{CHECK} Successfully generated: {output_file}")
            sys.exit(0)
        else:
            print(f"{WARNING} Validation warnings - please review the output")
            sys.exit(1)
            
    except ImportError as e:
        print()
        print("ERROR: Missing required library")
        print("-" * 70)
        print(f"  {str(e)}")
        print()
        print("Please install openpyxl:")
        print("  pip install openpyxl")
        print()
        sys.exit(1)
        
    except Exception as e:
        print()
        print("ERROR: Generation failed")
        print("-" * 70)
        print(f"  {str(e)}")
        print()
        import traceback
        traceback.print_exc()
        sys.exit(1)


# =============================================================================
# END OF GENERATOR SCRIPT
# =============================================================================
#
# Document Control:
#   Version: 1.0
#   Created: 2025-01-11
#   Author: ISMS Implementation Team
#
# Change History:
#   1.0 - Initial version
#       - 11 sheets for comprehensive privileged access assessment
#       - 150-row privileged account inventory with tier classification
#       - Admin Tiering Model (Tier 0/1/2) implementation guidance
#       - PAM vault coverage tracking (100% target for shared accounts)
#       - Hardware token deployment for Tier 0
#       - Credential rotation and quarterly review tracking
#       - Tier isolation violation monitoring
#
# Dependencies:
#   - Python 3.7+
#   - openpyxl >= 3.0.0
#
# Output:
#   Privileged_Accounts_Assessment_YYYYMMDD.xlsx
#
# =============================================================================
