#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.9.2 - Change Control Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.9: Configuration Management
Assessment Domain 2 of 4: Change Control and Configuration Updates

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific change management processes, CAB operations, and
approval workflows.

Key customization areas:
1. Change classification criteria (match your change management framework)
2. CAB membership and approval authorities (align with organizational roles)
3. Approval workflow tiers (adapt to your governance structure)
4. Testing requirements and environments (specific to your infrastructure)
5. Change management system integration (ServiceNow, Jira, etc.)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.9 Configuration Management Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
change control and configuration update processes against ISO 27001:2022
Control A.8.9 requirements.

**Purpose:**
Enables systematic assessment of change classification, CAB operations, approval
workflows, testing procedures, and rollback capabilities to ensure controlled
and authorized configuration modifications.

**Assessment Scope:**
- Change classification framework (Standard, Normal, Emergency)
- Change Advisory Board (CAB) operations and governance
- Approval workflow execution and authorization
- Testing and validation requirements compliance
- Rollback procedure readiness and testing
- Change success metrics and KPI tracking
- Post-implementation review completion
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and change control standards
2. Change Classification - Change type inventory and classification criteria
3. CAB Operations - CAB membership, meeting schedule, decision tracking
4. Approval Workflows - Approval chain execution and authorization evidence
5. Testing & Validation - Test planning, execution, and results
6. Rollback Procedures - Rollback readiness and testing evidence
7. Change Metrics - Success rate, emergency changes, SLA compliance
8. Gap Analysis - Non-compliant changes and remediation requirements
9. Evidence Register - Audit evidence tracking (100+ entries)
10. Approval & Sign-Off - Three-tier approval workflow

**Key Features:**
- Data validation with change type and risk level dropdowns
- Conditional formatting for approval status and SLA compliance
- Automated gap identification for untested or unapproved changes
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with change management systems

**Integration:**
This assessment feeds into the A.8.9.5 Compliance Dashboard, which
consolidates data from all four configuration management assessment domains
for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a89_2_change_control.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a89_2_change_control.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a89_2_change_control.py --date 20250127

Output:
    File: ISMS_IMP_A_8_9_2_Change_Control_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize change classification criteria
    2. Validate CAB membership and meeting schedule
    3. Document approval workflows by change risk level
    4. Assess testing and validation compliance
    5. Verify rollback procedure readiness
    6. Calculate change success metrics
    7. Conduct gap analysis for non-compliant changes
    8. Define remediation actions with timelines
    9. Collect and link audit evidence (change requests, CAB minutes)
    10. Obtain three-tier stakeholder approvals
    11. Feed results into A.8.9.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.9
Assessment Domain:    2 of 4 (Change Control and Configuration Updates)
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.9: Configuration Management Policy (Consolidated)
    - ISMS-POL-A.8.9, Section 2.3: Change Control & Configuration Updates
    - ISMS-CTX-A.8.9: Configuration Management Reference (NOT ISMS)
    - ISMS-CTX-A.8.9, Part 2: Change Management Implementation Guide
    - ISMS-CTX-A.8.9-Evidence-Collection: Evidence Collection Guide
    - ISMS-IMP-A.8.9.1: Baseline Configuration Assessment (Domain 1)
    - ISMS-IMP-A.8.9.3: Configuration Monitoring Assessment (Domain 3)
    - ISMS-IMP-A.8.9.4: Security Hardening Assessment (Domain 4)
    - ISMS-IMP-A.8.9.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.9.2 specification
    - Supports comprehensive change control evaluation
    - Integrated with A.8.9.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Change Management Standards:**
Change management best practices evolve. Review ITIL 4 change enablement,
organizational change success metrics, and industry benchmarks quarterly.
Emergency change abuse and inadequate testing must be identified and corrected.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of CAB operations, approvals, and testing.

**Data Protection:**
Assessment workbooks contain sensitive operational details including:
- Change request records with system details
- CAB meeting minutes with decision rationale
- Testing results and vulnerability information
- Rollback procedures and recovery capabilities

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Monthly: Check change success metrics and emergency change rate
- Quarterly: Review CAB effectiveness and approval SLA compliance
- Annually: Complete reassessment of change control processes
- Ad-hoc: When change management process changes or incidents occur

**Quality Assurance:**
Have change management coordinators and CAB leadership validate assessments
before using results for compliance reporting or process improvement decisions.

**Regulatory Alignment:**
Ensure change control practices align with applicable requirements:
- ISO 27001:2022: Control A.8.9 change management
- ITIL 4: Change enablement practices
- Sector-specific: Regulatory change control requirements
- Internal: Organizational change management policies

Customize assessment criteria to include regulatory-specific requirements.

================================================================================
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •

from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

# ============================================================================
# CONFIGURATION SECTION - CUSTOMIZE FOR YOUR ENVIRONMENT
# ============================================================================

# File output configuration
FILENAME = f"ISMS-IMP-A.8.X.X_..._Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"

# Workbook metadata
WORKBOOK_TITLE = "Change Control Assessment"
WORKBOOK_VERSION = "1.0"
DOCUMENT_ID = "ISMS-IMP-A.8.9.2"

# CUSTOMIZE: Change management dropdown values
CHANGE_TYPE = ["Standard", "Normal", "Emergency", "Hot Fix"]
CHANGE_PRIORITY = ["🔴 P1-Critical", "🟡 P2-High", "🟢 P3-Medium", "⭕ P4-Low"]
CHANGE_STATUS = ["📝 Draft", "📤 Submitted", "\u2705 Approved", "🧪 In Testing", "📅 Scheduled", 
                "⏳ Implementing", "\u2705 Completed", "\u274C Failed", "🔄 Rolled Back", "\u274C Cancelled"]

APPROVAL_TIER = ["Single-Tier", "Two-Tier", "Three-Tier", "Emergency"]
APPROVAL_STATUS_TIER = ["⏳ Pending", "\u2705 Approved", "\u274C Rejected", "➖ N/A"]
APPROVAL_METHOD = ["CAB Meeting", "Email Approval", "Emergency Verbal", "Automated (Standard)", "Not Applicable"]

USER_IMPACT = ["None", "Minimal", "Moderate", "Significant", "Severe"]
SERVICE_DOWNTIME = ["None", "<1 hour", "1-4 hours", "4-8 hours", ">8 hours"]
RISK_LEVEL = ["⭕ Low", "🟢 Medium", "🟡 High", "🔴 Critical"]
YES_NO = ["Yes", "No"]
YES_NO_NA = ["Yes", "No", "N/A"]
YES_NO_PARTIAL_NA = ["Yes", "No", "Partially", "N/A"]

TEST_ENVIRONMENT = ["Dev", "Test", "Staging", "UAT", "Production (Non-Critical)", "None"]
TESTING_STATUS = ["🔴 Not Started", "⏳ In Progress", "\u2705 Completed", "\u274C Failed", "⚡ Abbreviated (Emergency)"]
GO_NOGO = ["Go", "No-Go", "Go with Conditions", "N/A"]

IMPLEMENTATION_METHOD = ["Manual", "Automated Script", "Semi-Automated", "Assisted (Vendor)"]
VERIFICATION_STATUS = ["\u2705 Successful", "\u26A0\uFE0F Partial Success", "\u274C Failed", "❓ Not Verified"]
ISSUES_RESOLVED = ["Yes", "No", "Partially", "N/A"]
IMPLEMENTATION_STATUS = ["\u2705 Successful", "\u274C Failed", "🔄 Rolled Back", "⏳ In Progress"]

ROLLBACK_TEST_RESULTS = ["\u2705 Successful", "\u274C Failed", "\u26A0\uFE0F Partially Successful", "❓ Not Tested"]
DATA_LOSS_RISK = ["None", "Minimal", "Moderate", "Significant"]
ROLLBACK_APPROVAL = ["Yes (same as forward)", "Yes (expedited)", "No (automatic)"]

EMERGENCY_TYPE = ["Security Incident", "Service Outage", "Critical Bug", "Vulnerability Remediation", "Other"]
EMERGENCY_APPROVAL_METHOD = ["Verbal (CIO/CISO)", "Email (Expedited)", "CAB Chair Authorization"]
POST_IMPL_DOC = ["Yes", "No", "In Progress"]
CAB_REVIEW_OUTCOME = ["\u2705 Approved", "\u26A0\uFE0F Approved with Remediation", "\u274C Disapproved (requires reversal)", "⏳ Not Yet Reviewed"]
JUSTIFICATION_VALID = ["Yes", "No", "Questionable"]

EVIDENCE_TYPE = ["Approval Record", "Test Results", "Implementation Log", "Rollback Test", 
                "CAB Minutes", "Email Approval", "Change Request", "Other"]
EVIDENCE_CLASSIFICATION = ["Public", "Internal", "Confidential", "Restricted"]
RETENTION_PERIOD = ["1 Year", "3 Years", "5 Years", "7 Years", "Indefinite"]
EVIDENCE_VERIFICATION = ["\u2705 Verified", "🔍 Needs Verification", "\u274C Missing", "⏰ Outdated"]

APPROVAL_DECISION = ["Approved", "Approved with Conditions", "Not Approved - Revisions Required"]

# CUSTOMIZE: Color scheme
COLORS = {
    'header_main': '003366',          # Dark Blue
    'header_sub': '4472C4',           # Blue
    'column_header': 'D9D9D9',        # Light Gray
    'input_cell': 'FFFFFF',           # White
    'protected_cell': 'F2F2F2',       # Very Light Gray
    'compliant': 'C6EFCE',            # Green
    'partial': 'FFEB9C',              # Yellow
    'non_compliant': 'FFC7CE',        # Red
    'excluded': 'D9D9D9',             # Gray
    'critical': 'C00000',             # Dark Red
    'info_bg': 'E7E6E6',              # Light Gray
    'light_green': 'E2EFDA',          # Light Green
    'orange': 'FFA500'                # Orange
}

# ============================================================================
# STYLE DEFINITIONS
# ============================================================================

def create_styles():
    """Creates and returns a dictionary of reusable styles."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    styles = {
        'header_main': {
            'font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_main'], 
                              end_color=COLORS['header_main'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'header_sub': {
            'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_sub'], 
                              end_color=COLORS['header_sub'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'column_header': {
            'font': Font(name='Calibri', size=11, bold=True),
            'fill': PatternFill(start_color=COLORS['column_header'], 
                              end_color=COLORS['column_header'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'data_cell': {
            'font': Font(name='Calibri', size=11),
            'fill': PatternFill(start_color=COLORS['input_cell'], 
                              end_color=COLORS['input_cell'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=False),
            'border': thin_border
        },
        'protected_cell': {
            'font': Font(name='Calibri', size=11, italic=True),
            'fill': PatternFill(start_color=COLORS['protected_cell'], 
                              end_color=COLORS['protected_cell'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top'),
            'border': thin_border
        },
        'info_text': {
            'font': Font(name='Calibri', size=11),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True)
        },
        'section_header': {
            'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_sub'], 
                              end_color=COLORS['header_sub'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': thin_border
        }
    }
    return styles

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

def set_column_widths(ws, widths):
    """Set column widths for a worksheet."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def create_data_validation(values, allow_blank=True):
    """Create a data validation object for dropdowns."""
    formula = f'"{",".join(values)}"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = 'Invalid entry'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select from the dropdown'
    dv.promptTitle = 'Selection Required'
    return dv

def protect_formula_cells(ws, start_row, end_row, formula_columns):
    """Protect cells containing formulas."""
    for row in range(start_row, end_row + 1):
        for col in formula_columns:
            ws[f'{col}{row}'].protection = Protection(locked=True)

# ============================================================================
# SHEET CREATION FUNCTIONS
# ============================================================================

def create_instructions_sheet(wb, styles):
    """Create the Instructions and Legend sheet."""
    ws = wb.create_sheet("Instructions", 0)
    ws.column_dimensions['A'].width = 100
    
    # Title
    ws.merge_cells('A1:A2')
    ws['A1'] = "ISMS Control A.8.9 - Change Control Assessment"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color=COLORS['header_main'], 
                                end_color=COLORS['header_main'], 
                                fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Document metadata
    ws['A3'] = "Document ID:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = DOCUMENT_ID
    
    ws['A4'] = "Assessment:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "Change Control Assessment"
    
    ws['A5'] = "Version:"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = "1.0"
    
    ws['A6'] = "Generated:"
    ws['A6'].font = Font(bold=True)
    ws['B6'] = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    ws.column_dimensions['B'].width = 40
    
    # CUSTOMIZE: Update instructions for your organization
    instructions = [
        "",
        "ASSESSMENT OVERVIEW",
        "━" * 80,
        "",
        "Purpose:",
        "This workbook evaluates the effectiveness of configuration change management processes. "
        "It verifies that all changes follow controlled procedures including impact assessment, approval workflows, "
        "testing validation, and rollback capability. Provides evidence of ISO 27001:2022 Control A.8.9 compliance.",
        "",
        "Assessment Scope:",
        "\u2022 All configuration changes (Standard, Normal, Emergency, Hot Fix)",
        "\u2022 Multi-tier approval workflows and authorization chains",
        "\u2022 Pre-deployment testing and validation",
        "\u2022 Change implementation and post-implementation verification",
        "\u2022 Rollback capability and procedures",
        "\u2022 Emergency change management and retrospective review",
        "",
        "CHANGE TYPES AND PRIORITIES",
        "━" * 80,
        "",
        "Change Types:",
        "\u2022 Standard: Pre-approved, low-risk changes (routine patching, scheduled maintenance)",
        "\u2022 Normal: Planned changes requiring individual assessment and CAB approval",
        "\u2022 Emergency: Urgent changes for critical incidents (expedited approval, post-review required)",
        "\u2022 Hot Fix: Immediate changes for service restoration (minimal approval, full retrospective)",
        "",
        "Priorities:",
        "\u2022 P1-Critical: Immediate action (<4 hours) - typically Hot Fix or Emergency",
        "\u2022 P2-High: Urgent (<24 hours) - Emergency or Normal",
        "\u2022 P3-Medium: Important but scheduled (<7 days) - Normal or Standard",
        "\u2022 P4-Low: Minor, can be bundled (<30 days) - Standard or Normal",
        "",
        "Approval Tiers (Risk-Based):",
        "\u2022 Single-Tier: Standard changes, low-risk (Change Coordinator approval)",
        "\u2022 Two-Tier: Normal production changes (Technical Lead + Service Owner)",
        "\u2022 Three-Tier: Critical systems, high-risk (Technical Lead + Service Owner + CAB/CIO/CISO)",
        "\u2022 Emergency: Expedited (minimum 2 of: CIO/CISO/IT Manager, CAB review within 5 days)",
        "",
        "WHO SHOULD COMPLETE THIS ASSESSMENT",
        "━" * 80,
        "",
        "Change Coordinator (Preparer):",
        "\u2022 Document all configuration changes in Change_Request_Register",
        "\u2022 Complete Impact_Assessment before submission",
        "\u2022 Track approval workflow and testing",
        "\u2022 Log implementation and outcomes",
        "\u2022 Maintain ongoing change records",
        "",
        "Configuration Manager (Reviewer):",
        "\u2022 Quarterly review of change management effectiveness",
        "\u2022 Verify process compliance (approval, testing, documentation)",
        "\u2022 Analyze success metrics and trends",
        "\u2022 Identify process improvement opportunities",
        "\u2022 Review emergency changes for legitimacy",
        "",
        "IT Manager / CISO (Approver):",
        "\u2022 Review overall change management health",
        "\u2022 Approve assessment findings and remediation plans",
        "\u2022 Authorize process improvements",
        "",
        "COMPLETION INSTRUCTIONS",
        "━" * 80,
        "",
        "Ongoing Change Tracking:",
        "1. For each configuration change:",
        "   \u2022 Create entry in Change_Request_Register (assign unique Change ID)",
        "   \u2022 Complete Impact_Assessment before approval submission",
        "   \u2022 Document approval in Change_Approval_Workflow",
        "   \u2022 Record testing in Testing_Validation",
        "   \u2022 Document rollback plan in Rollback_Capability (high-risk changes)",
        "   \u2022 Log implementation in Implementation_Log",
        "   \u2022 Update Change_Request_Register status throughout lifecycle",
        "",
        "2. For emergency changes:",
        "   \u2022 Follow expedited approval process",
        "   \u2022 Document in Emergency_Changes sheet immediately after implementation",
        "   \u2022 Complete post-implementation documentation within 24 hours",
        "   \u2022 Schedule CAB retrospective review within 5 business days",
        "",
        "Monthly Review (Change Coordinator):",
        "1. Review previous month's changes for completeness",
        "2. Verify all documentation is present",
        "3. Follow up on incomplete records",
        "4. Review Change_Success_Metrics for trends",
        "5. Document gaps for quarterly assessment",
        "",
        "Quarterly Assessment (Configuration Manager):",
        "1. Comprehensive review of all sheets",
        "2. Analyze Compliance_Dashboard metrics",
        "3. Review emergency change ratio and justifications",
        "4. Identify process improvement opportunities",
        "5. Compile evidence and complete Reviewer Sign-Off",
        "",
        "Annual/Semi-Annual Approval (IT Manager/CISO):",
        "1. Review change management effectiveness",
        "2. Approve remediation plans",
        "3. Complete Approver Sign-Off",
        "",
        "WORKBOOK STRUCTURE",
        "━" * 80,
        "",
        "This workbook contains 12 sheets:",
        "",
        "1.  Instructions - This sheet (usage guidance)",
        "2.  Change_Request_Register - All configuration changes (100 rows)",
        "3.  Change_Approval_Workflow - Approval chain tracking (100 rows)",
        "4.  Impact_Assessment - Risk and impact analysis (100 rows)",
        "5.  Testing_Validation - Pre-deployment testing records (100 rows)",
        "6.  Implementation_Log - Change execution records (100 rows)",
        "7.  Rollback_Capability - Rollback procedures and testing (100 rows)",
        "8.  Emergency_Changes - Emergency change tracking (50 rows)",
        "9.  Change_Success_Metrics - Auto-calculated success rates (dashboard)",
        "10. Compliance_Dashboard - Process adherence metrics (dashboard)",
        "11. Evidence_Register - Supporting evidence (100 rows)",
        "12. Approval_Sign_Off - Three-tier approval signatures",
        "",
        "LEGEND - STATUS VALUES AND COLOR CODING",
        "━" * 80,
        "",
        "Change Status:",
        "  🟢 Completed (Green) - Change successfully implemented and verified",
        "  🟡 In Progress (Yellow) - Change in testing, scheduled, or implementing",
        "  🔴 Failed (Red) - Change implementation failed",
        "  🟠 Rolled Back (Orange) - Change was reversed due to issues",
        "",
        "Priority:",
        "  🔴 P1-Critical (Red) - Immediate action required",
        "  🟡 P2-High (Yellow) - Urgent, within 24 hours",
        "  ⚪ P3-Medium, P4-Low (Normal) - Scheduled implementation",
        "",
        "Approval Status:",
        "  🟢 Approved (Green) - Approved by all required tiers",
        "  🟡 Pending (Yellow) - Awaiting approval",
        "  🔴 Rejected (Red) - Change rejected, requires rework",
        "",
        "Testing Results:",
        "  🟢 ≥95% Pass Rate (Green) - Adequate testing",
        "  🟡 80-94% Pass Rate (Yellow) - Marginal, review issues",
        "  🔴 <80% Pass Rate (Red) - Insufficient quality, do not proceed",
        "",
        "Implementation Status:",
        "  🟢 Successful (Green) - Change achieved intended result",
        "  🔴 Failed (Red) - Change did not work as planned",
        "  🟠 Rolled Back (Orange) - Change was reversed",
        "",
        "COMPLIANCE TARGETS",
        "━" * 80,
        "",
        "Change Success Rate: ≥95%",
        "Emergency Change Ratio: <10% of total changes",
        "Approval Compliance: 100% (all changes have proper approval)",
        "Testing Coverage: 100% for Normal changes, 80% for Emergency",
        "Rollback Readiness: 100% for High/Critical risk changes",
        "Overall Compliance: ≥95%",
        "",
        "IMPORTANT NOTES",
        "━" * 80,
        "",
        "\u2022 Protected cells (gray background) contain formulas - do not edit",
        "\u2022 Use dropdowns for standardized entries",
        "\u2022 Date fields: DD.MM.YYYY format, DateTime: DD.MM.YYYY HH:MM",
        "\u2022 Change_Success_Metrics and Compliance_Dashboard auto-update",
        "\u2022 Emergency changes require CAB retrospective review within 5 business days",
        "\u2022 Standard changes are pre-approved but still require documentation",
        "\u2022 All high-risk changes require documented rollback procedures",
        "",
        "SUPPORT AND QUESTIONS",
        "━" * 80,
        "",
        "Configuration Manager: [contact information]",
        "Change Advisory Board: [contact information]",
        "ISMS Team: [contact information]",
        "",
        "For technical issues: [IT support]",
        "For policy questions: Reference ISMS-POL-A.8.9-S2.2",
        "",
        "━" * 80,
        f"Generated: {datetime.now().strftime('%d.%m.%Y')} | Version: {WORKBOOK_VERSION} | Document ID: {DOCUMENT_ID}",
    ]
    
    row = 8
    for line in instructions:
        ws[f'A{row}'] = line
        if line.startswith("━"):
            ws[f'A{row}'].font = Font(name='Calibri', size=11, color='666666')
        elif line.isupper() and len(line) > 5 and not line.startswith("  "):
            apply_style(ws[f'A{row}'], styles['section_header'])
            ws.row_dimensions[row].height = 20
        else:
            apply_style(ws[f'A{row}'], styles['info_text'])
        row += 1
    
    return ws

def create_change_request_register_sheet(wb, styles):
    """Create Sheet 2: Change_Request_Register"""
    ws = wb.create_sheet("Change_Request_Register")
    
    # Set column widths
    widths = {
        'A': 18,  # Change ID
        'B': 35,  # Change Title
        'C': 15,  # Change Type
        'D': 15,  # Priority
        'E': 35,  # Affected Systems/Assets
        'F': 20,  # Requestor Name
        'G': 20,  # Requestor Contact
        'H': 15,  # Request Date
        'I': 15,  # Required Implementation Date
        'J': 18,  # Change Status
        'K': 18,  # Current Phase
        'L': 12,  # Days in Current Phase
        'M': 18,  # Overall Status Indicator
        'N': 40   # Notes
    }
    set_column_widths(ws, widths)
    
    # Title row
    ws.merge_cells('A1:N1')
    ws['A1'] = "Change Request Register - Configuration Change Tracking"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    # Column headers
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Change Type', 'D': 'Priority',
        'E': 'Affected Systems/Assets', 'F': 'Requestor Name', 'G': 'Requestor Contact',
        'H': 'Request Date', 'I': 'Required Implementation Date', 'J': 'Change Status',
        'K': 'Current Phase', 'L': 'Days in Current Phase', 'M': 'Overall Status Indicator',
        'N': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    # Create 100 data rows
    num_rows = 100
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['K', 'L', 'M']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Add formulas
    for row in range(3, 3 + num_rows):
        # Column K: Current Phase
        ws[f'K{row}'] = f'=IF(J{row}="Draft","Planning",IF(J{row}="Submitted","Approval",IF(OR(J{row}="Approved",J{row}="In Testing"),"Testing",IF(J{row}="Scheduled","Pre-Implementation",IF(J{row}="Implementing","Implementation",IF(OR(J{row}="Completed",J{row}="Failed",J{row}="Rolled Back",J{row}="Cancelled"),"Closed","Unknown"))))))'
        
        # Column L: Days in Current Phase
        ws[f'L{row}'] = f'=IF(H{row}="","",TODAY()-H{row})'
        
        # Column M: Overall Status Indicator
        ws[f'M{row}'] = f'=IF(OR(J{row}="Completed",J{row}="Cancelled"),"Complete",IF(L{row}>30,"Delayed",IF(L{row}>14,"At Risk","On Track")))'
    
    # Data validations
    change_type_dv = create_data_validation(CHANGE_TYPE, allow_blank=False)
    ws.add_data_validation(change_type_dv)
    change_type_dv.add(f'C3:C{2+num_rows}')
    
    priority_dv = create_data_validation(CHANGE_PRIORITY, allow_blank=False)
    ws.add_data_validation(priority_dv)
    priority_dv.add(f'D3:D{2+num_rows}')
    
    status_dv = create_data_validation(CHANGE_STATUS, allow_blank=False)
    ws.add_data_validation(status_dv)
    status_dv.add(f'J3:J{2+num_rows}')
    
    # Conditional formatting
    ws.conditional_formatting.add(f'C3:C{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Emergency"'], 
                   font=Font(color=COLORS['orange'])))
    ws.conditional_formatting.add(f'C3:C{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Hot Fix"'], 
                   font=Font(bold=True, color='9C0006')))
    
    ws.conditional_formatting.add(f'D3:D{2+num_rows}',
        CellIsRule(operator='equal', formula=['"P1-Critical"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'D3:D{2+num_rows}',
        CellIsRule(operator='equal', formula=['"P2-High"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'J3:J{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Completed"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'J3:J{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Failed"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'M3:M{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Complete"'], 
                   font=Font(bold=True, color='006100')))
    ws.conditional_formatting.add(f'M3:M{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Delayed"'], 
                   font=Font(bold=True, color='9C0006')))
    
    ws.freeze_panes = 'B3'
    protect_formula_cells(ws, 3, 2 + num_rows, ['K', 'L', 'M'])
    
    return ws

def create_change_approval_workflow_sheet(wb, styles):
    """Create Sheet 3: Change_Approval_Workflow"""
    ws = wb.create_sheet("Change_Approval_Workflow")
    
    widths = {
        'A': 18, 'B': 30, 'C': 18, 'D': 20, 'E': 18, 'F': 15,
        'G': 20, 'H': 18, 'I': 15, 'J': 20, 'K': 18, 'L': 15,
        'M': 25, 'N': 30, 'O': 18, 'P': 15, 'Q': 40
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:Q1')
    ws['A1'] = "Change Approval Workflow Tracking"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Approval Tier Required',
        'D': 'Tier 1 Approver Name', 'E': 'Tier 1 Approval Status', 'F': 'Tier 1 Approval Date',
        'G': 'Tier 2 Approver Name', 'H': 'Tier 2 Approval Status', 'I': 'Tier 2 Approval Date',
        'J': 'Tier 3 Approver Name', 'K': 'Tier 3 Approval Status', 'L': 'Tier 3 Approval Date',
        'M': 'Approval Method', 'N': 'Approval Reference', 'O': 'Overall Approval Status',
        'P': 'Days to Full Approval', 'Q': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    num_rows = 100
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['O', 'P']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Add complex formulas for approval status
    for row in range(3, 3 + num_rows):
        # Column O: Overall Approval Status (complex logic)
        ws[f'O{row}'] = f'=IF(C{row}="Single-Tier",IF(E{row}="Approved","Approved",IF(E{row}="Rejected","Rejected","Pending")),IF(C{row}="Two-Tier",IF(AND(E{row}="Approved",H{row}="Approved"),"Approved",IF(OR(E{row}="Rejected",H{row}="Rejected"),"Rejected","Pending")),IF(C{row}="Three-Tier",IF(AND(E{row}="Approved",H{row}="Approved",K{row}="Approved"),"Approved",IF(OR(E{row}="Rejected",H{row}="Rejected",K{row}="Rejected"),"Rejected","Pending")),"Unknown")))'
        
        # Column P: Days to Full Approval (simplified - would need VLOOKUP to Change_Request_Register in real implementation)
        ws[f'P{row}'] = f'=IF(O{row}="Approved",IF(C{row}="Single-Tier",IF(F{row}="","",F{row}-TODAY()+365),IF(C{row}="Two-Tier",IF(I{row}="","",I{row}-TODAY()+365),IF(L{row}="","",L{row}-TODAY()+365))),"")'
    
    # Data validations
    tier_dv = create_data_validation(APPROVAL_TIER, allow_blank=False)
    ws.add_data_validation(tier_dv)
    tier_dv.add(f'C3:C{2+num_rows}')
    
    for col in ['E', 'H', 'K']:
        status_dv = create_data_validation(APPROVAL_STATUS_TIER, allow_blank=True)
        ws.add_data_validation(status_dv)
        status_dv.add(f'{col}3:{col}{2+num_rows}')
    
    method_dv = create_data_validation(APPROVAL_METHOD, allow_blank=True)
    ws.add_data_validation(method_dv)
    method_dv.add(f'M3:M{2+num_rows}')
    
    # Conditional formatting
    for col in ['E', 'H', 'K']:
        ws.conditional_formatting.add(f'{col}3:{col}{2+num_rows}',
            CellIsRule(operator='equal', formula=['"Approved"'], 
                       fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
        ws.conditional_formatting.add(f'{col}3:{col}{2+num_rows}',
            CellIsRule(operator='equal', formula=['"Rejected"'], 
                       fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.freeze_panes = 'B3'
    protect_formula_cells(ws, 3, 2 + num_rows, ['O', 'P'])
    
    return ws

def create_impact_assessment_sheet(wb, styles):
    """Create Sheet 4: Impact_Assessment"""
    ws = wb.create_sheet("Impact_Assessment")
    
    widths = {
        'A': 18, 'B': 30, 'C': 12, 'D': 35, 'E': 15, 'F': 15,
        'G': 20, 'H': 15, 'I': 35, 'J': 35, 'K': 15, 'L': 20,
        'M': 35, 'N': 35, 'O': 12, 'P': 25, 'Q': 15
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:Q1')
    ws['A1'] = "Impact Assessment - Risk Analysis per Change"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Affected Systems Count',
        'D': 'Affected Systems Detail', 'E': 'User Impact', 'F': 'User Count Affected',
        'G': 'Service Downtime Required', 'H': 'Risk Level', 'I': 'Risk Description',
        'J': 'Mitigation Strategies', 'K': 'Rollback Required', 'L': 'Estimated Rollback Time',
        'M': 'Dependencies', 'N': 'Business Justification', 'O': 'Risk Score',
        'P': 'Assessment Completed By', 'Q': 'Assessment Date'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    num_rows = 100
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col == 'O':
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Add Risk Score formula
    for row in range(3, 3 + num_rows):
        ws[f'O{row}'] = f'=IF(H{row}="Low",1,IF(H{row}="Medium",2,IF(H{row}="High",3,IF(H{row}="Critical",4,0))))*IF(E{row}="None",1,IF(E{row}="Minimal",2,IF(E{row}="Moderate",3,IF(E{row}="Significant",4,IF(E{row}="Severe",5,0)))))'
    
    # Data validations
    user_impact_dv = create_data_validation(USER_IMPACT, allow_blank=False)
    ws.add_data_validation(user_impact_dv)
    user_impact_dv.add(f'E3:E{2+num_rows}')
    
    downtime_dv = create_data_validation(SERVICE_DOWNTIME, allow_blank=False)
    ws.add_data_validation(downtime_dv)
    downtime_dv.add(f'G3:G{2+num_rows}')
    
    risk_dv = create_data_validation(RISK_LEVEL, allow_blank=False)
    ws.add_data_validation(risk_dv)
    risk_dv.add(f'H3:H{2+num_rows}')
    
    rollback_dv = create_data_validation(YES_NO, allow_blank=False)
    ws.add_data_validation(rollback_dv)
    rollback_dv.add(f'K3:K{2+num_rows}')
    
    # Conditional formatting
    ws.conditional_formatting.add(f'H3:H{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Critical"'], 
                   fill=PatternFill(start_color=COLORS['critical'], end_color=COLORS['critical'], fill_type='solid'),
                   font=Font(bold=True, color='FFFFFF')))
    ws.conditional_formatting.add(f'H3:H{2+num_rows}',
        CellIsRule(operator='equal', formula=['"High"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'O3:O{2+num_rows}',
        CellIsRule(operator='greaterThanOrEqual', formula=['12'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'O3:O{2+num_rows}',
        CellIsRule(operator='between', formula=['6', '11'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    
    ws.freeze_panes = 'B3'
    protect_formula_cells(ws, 3, 2 + num_rows, ['O'])
    
    return ws

def create_testing_validation_sheet(wb, styles):
    """Create Sheet 5: Testing_Validation"""
    ws = wb.create_sheet("Testing_Validation")
    
    widths = {
        'A': 18, 'B': 30, 'C': 15, 'D': 20, 'E': 15, 'F': 15,
        'G': 12, 'H': 30, 'I': 12, 'J': 12, 'K': 12, 'L': 12,
        'M': 12, 'N': 12, 'O': 18, 'P': 18, 'Q': 20, 'R': 15, 'S': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:S1')
    ws['A1'] = "Testing Validation Records"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Testing Required',
        'D': 'Test Environment', 'E': 'Test Start Date', 'F': 'Test End Date',
        'G': 'Test Duration (Days)', 'H': 'Test Plan Reference', 'I': 'Test Cases Executed',
        'J': 'Test Cases Passed', 'K': 'Test Cases Failed', 'L': 'Test Pass Rate %',
        'M': 'Critical Issues Found', 'N': 'Issues Resolved Before Deployment',
        'O': 'Testing Status', 'P': 'Go/No-Go Decision', 'Q': 'Decision Maker',
        'R': 'Decision Date', 'S': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    num_rows = 100
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['G', 'L']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Formulas
    for row in range(3, 3 + num_rows):
        ws[f'G{row}'] = f'=IF(OR(E{row}="",F{row}=""),"",F{row}-E{row})'
        ws[f'L{row}'] = f'=IF(I{row}=0,"",J{row}/I{row}*100)'
        ws[f'L{row}'].number_format = '0.0'
    
    # Data validations
    test_req_dv = create_data_validation(YES_NO_NA, allow_blank=False)
    ws.add_data_validation(test_req_dv)
    test_req_dv.add(f'C3:C{2+num_rows}')
    
    env_dv = create_data_validation(TEST_ENVIRONMENT, allow_blank=True)
    ws.add_data_validation(env_dv)
    env_dv.add(f'D3:D{2+num_rows}')
    
    status_dv = create_data_validation(TESTING_STATUS, allow_blank=False)
    ws.add_data_validation(status_dv)
    status_dv.add(f'O3:O{2+num_rows}')
    
    gonogo_dv = create_data_validation(GO_NOGO, allow_blank=True)
    ws.add_data_validation(gonogo_dv)
    gonogo_dv.add(f'P3:P{2+num_rows}')
    
    # Conditional formatting
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='greaterThanOrEqual', formula=['95'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='between', formula=['80', '94.9'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='lessThan', formula=['80'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Go"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"No-Go"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    
    ws.freeze_panes = 'B3'
    protect_formula_cells(ws, 3, 2 + num_rows, ['G', 'L'])
    
    return ws

def create_implementation_log_sheet(wb, styles):
    """Create Sheet 6: Implementation_Log"""
    ws = wb.create_sheet("Implementation_Log")
    
    widths = {
        'A': 18, 'B': 30, 'C': 20, 'D': 20, 'E': 20, 'F': 12,
        'G': 20, 'H': 20, 'I': 40, 'J': 35, 'K': 20, 'L': 30,
        'M': 35, 'N': 18, 'O': 18, 'P': 18, 'Q': 15, 'R': 20, 'S': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:S1')
    ws['A1'] = "Implementation Log - Change Execution Records"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Scheduled Implementation Date/Time',
        'D': 'Actual Implementation Date/Time', 'E': 'Implementation Completed Date/Time',
        'F': 'Implementation Duration (Hours)', 'G': 'Implementer Name',
        'H': 'Implementation Method', 'I': 'Implementation Steps Performed',
        'J': 'Deviations from Plan', 'K': 'Post-Implementation Verification',
        'L': 'Verification Method', 'M': 'Issues Encountered', 'N': 'Issues Resolved',
        'O': 'Implementation Status', 'P': 'Change Outcome', 'Q': 'Rollback Triggered',
        'R': 'Rollback Completion Time', 'S': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    num_rows = 100
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['F', 'P']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Formulas
    for row in range(3, 3 + num_rows):
        ws[f'F{row}'] = f'=IF(OR(D{row}="",E{row}=""),"",ROUND((E{row}-D{row})*24,2))'
        ws[f'P{row}'] = f'=IF(O{row}="Successful","Success",IF(O{row}="Failed","Failure",IF(O{row}="Rolled Back","Rollback Required","Unknown")))'
    
    # Data validations
    method_dv = create_data_validation(IMPLEMENTATION_METHOD, allow_blank=False)
    ws.add_data_validation(method_dv)
    method_dv.add(f'H3:H{2+num_rows}')
    
    verif_dv = create_data_validation(VERIFICATION_STATUS, allow_blank=False)
    ws.add_data_validation(verif_dv)
    verif_dv.add(f'K3:K{2+num_rows}')
    
    resolved_dv = create_data_validation(ISSUES_RESOLVED, allow_blank=True)
    ws.add_data_validation(resolved_dv)
    resolved_dv.add(f'N3:N{2+num_rows}')
    
    impl_status_dv = create_data_validation(IMPLEMENTATION_STATUS, allow_blank=False)
    ws.add_data_validation(impl_status_dv)
    impl_status_dv.add(f'O3:O{2+num_rows}')
    
    rollback_dv = create_data_validation(YES_NO, allow_blank=False)
    ws.add_data_validation(rollback_dv)
    rollback_dv.add(f'Q3:Q{2+num_rows}')
    
    # Conditional formatting
    ws.conditional_formatting.add(f'O3:O{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Successful"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'O3:O{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Failed"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    
    ws.conditional_formatting.add(f'Q3:Q{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Yes"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.freeze_panes = 'B3'
    protect_formula_cells(ws, 3, 2 + num_rows, ['F', 'P'])
    
    return ws

def create_rollback_capability_sheet(wb, styles):
    """Create Sheet 7: Rollback_Capability"""
    ws = wb.create_sheet("Rollback_Capability")
    
    widths = {
        'A': 18, 'B': 30, 'C': 15, 'D': 18, 'E': 40, 'F': 35,
        'G': 20, 'H': 18, 'I': 15, 'J': 20, 'K': 35, 'L': 18,
        'M': 15, 'N': 25, 'O': 20, 'P': 18, 'Q': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:Q1')
    ws['A1'] = "Rollback Capability Assessment"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Rollback Required',
        'D': 'Rollback Procedure Documented', 'E': 'Rollback Document Location',
        'F': 'Rollback Trigger Criteria', 'G': 'Estimated Rollback Time',
        'H': 'Rollback Tested', 'I': 'Rollback Test Date', 'J': 'Rollback Test Results',
        'K': 'Rollback Dependencies', 'L': 'Data Loss Risk on Rollback',
        'M': 'Data Backup Verified', 'N': 'Rollback Approval Required',
        'O': 'Rollback Owner', 'P': 'Rollback Readiness', 'Q': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    num_rows = 100
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col == 'P':
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Formula for Rollback Readiness
    for row in range(3, 3 + num_rows):
        ws[f'P{row}'] = f'=IF(C{row}="No","N/A",IF(AND(D{row}="Yes",H{row}="Yes",J{row}="Successful",M{row}="Yes"),"Ready",IF(AND(D{row}="Yes",OR(H{row}="No",J{row}="Not Tested")),"Not Ready","Partially Ready")))'
    
    # Data validations
    rollback_req_dv = create_data_validation(YES_NO, allow_blank=False)
    ws.add_data_validation(rollback_req_dv)
    rollback_req_dv.add(f'C3:C{2+num_rows}')
    
    doc_dv = create_data_validation(YES_NO_NA, allow_blank=False)
    ws.add_data_validation(doc_dv)
    doc_dv.add(f'D3:D{2+num_rows}')
    
    tested_dv = create_data_validation(YES_NO_PARTIAL_NA, allow_blank=False)
    ws.add_data_validation(tested_dv)
    tested_dv.add(f'H3:H{2+num_rows}')
    
    results_dv = create_data_validation(ROLLBACK_TEST_RESULTS, allow_blank=True)
    ws.add_data_validation(results_dv)
    results_dv.add(f'J3:J{2+num_rows}')
    
    data_loss_dv = create_data_validation(DATA_LOSS_RISK, allow_blank=True)
    ws.add_data_validation(data_loss_dv)
    data_loss_dv.add(f'L3:L{2+num_rows}')
    
    backup_dv = create_data_validation(YES_NO_NA, allow_blank=False)
    ws.add_data_validation(backup_dv)
    backup_dv.add(f'M3:M{2+num_rows}')
    
    approval_dv = create_data_validation(ROLLBACK_APPROVAL, allow_blank=True)
    ws.add_data_validation(approval_dv)
    approval_dv.add(f'N3:N{2+num_rows}')
    
    # Conditional formatting
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Ready"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Not Ready"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Partially Ready"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'J3:J{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Successful"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'J3:J{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Failed"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'J3:J{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Not Tested"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Significant"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Moderate"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    
    ws.freeze_panes = 'B3'
    protect_formula_cells(ws, 3, 2 + num_rows, ['P'])
    
    return ws

def create_emergency_changes_sheet(wb, styles):
    """Create Sheet 8: Emergency_Changes"""
    ws = wb.create_sheet("Emergency_Changes")
    
    widths = {
        'A': 18, 'B': 30, 'C': 25, 'D': 40, 'E': 20, 'F': 20,
        'G': 20, 'H': 12, 'I': 25, 'J': 30, 'K': 20, 'L': 15,
        'M': 25, 'N': 18, 'O': 18, 'P': 40, 'Q': 40, 'R': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:R1')
    ws['A1'] = "Emergency Changes - Expedited Process Tracking"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    headers = {
        'A': 'Change ID', 'B': 'Change Title', 'C': 'Emergency Type',
        'D': 'Business Impact if Not Implemented', 'E': 'Emergency Declared By',
        'F': 'Emergency Declaration Time', 'G': 'Implementation Time',
        'H': 'Time to Implement (Hours)', 'I': 'Emergency Approval Method',
        'J': 'Emergency Approvers', 'K': 'Post-Implementation Documentation Completed',
        'L': 'CAB Retrospective Review Date', 'M': 'CAB Review Outcome',
        'N': 'Justification Valid', 'O': 'Process Abuse Indicator',
        'P': 'Lessons Learned', 'Q': 'Remediation Actions', 'R': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    num_rows = 50
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['H', 'O']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Formulas
    for row in range(3, 3 + num_rows):
        # Column H: Time to Implement (Hours)
        ws[f'H{row}'] = f'=IF(OR(F{row}="",G{row}=""),"",ROUND((G{row}-F{row})*24,2))'
        
        # Column O: Process Abuse Indicator
        ws[f'O{row}'] = f'=IF(AND(H{row}>48,N{row}="Questionable"),"Likely Abuse",IF(OR(H{row}>72,N{row}="No"),"Questionable","Legitimate"))'
    
    # Data validations
    emerg_type_dv = create_data_validation(EMERGENCY_TYPE, allow_blank=False)
    ws.add_data_validation(emerg_type_dv)
    emerg_type_dv.add(f'C3:C{2+num_rows}')
    
    emerg_method_dv = create_data_validation(EMERGENCY_APPROVAL_METHOD, allow_blank=False)
    ws.add_data_validation(emerg_method_dv)
    emerg_method_dv.add(f'I3:I{2+num_rows}')
    
    post_doc_dv = create_data_validation(POST_IMPL_DOC, allow_blank=False)
    ws.add_data_validation(post_doc_dv)
    post_doc_dv.add(f'K3:K{2+num_rows}')
    
    cab_review_dv = create_data_validation(CAB_REVIEW_OUTCOME, allow_blank=False)
    ws.add_data_validation(cab_review_dv)
    cab_review_dv.add(f'M3:M{2+num_rows}')
    
    justif_dv = create_data_validation(JUSTIFICATION_VALID, allow_blank=True)
    ws.add_data_validation(justif_dv)
    justif_dv.add(f'N3:N{2+num_rows}')
    
    # Conditional formatting
    ws.conditional_formatting.add(f'M3:M{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Approved"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'M3:M{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Disapproved (requires reversal)"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'M3:M{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Not Yet Reviewed"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'O3:O{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Legitimate"'], 
                   font=Font(color='006100')))
    ws.conditional_formatting.add(f'O3:O{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Questionable"'], 
                   font=Font(bold=True, color='9C6500')))
    ws.conditional_formatting.add(f'O3:O{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Likely Abuse"'], 
                   font=Font(bold=True, color='9C0006')))
    
    ws.conditional_formatting.add(f'K3:K{2+num_rows}',
        CellIsRule(operator='equal', formula=['"No"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'K3:K{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Yes"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    
    ws.freeze_panes = 'B3'
    protect_formula_cells(ws, 3, 2 + num_rows, ['H', 'O'])
    
    return ws

def create_change_success_metrics_sheet(wb, styles):
    """Create Sheet 9: Change_Success_Metrics (Dashboard)"""
    ws = wb.create_sheet("Change_Success_Metrics")
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "Change Success Metrics Dashboard"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    # SECTION A: Overall Change Metrics
    row = 3
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "OVERALL CHANGE METRICS"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    metrics = [
        ("Total Changes (All Types)", '=COUNTA(Change_Request_Register!A3:A102)-COUNTBLANK(Change_Request_Register!A3:A102)', "N/A", ""),
        ("Completed Changes", '=COUNTIF(Change_Request_Register!J3:J102,"Completed")', "N/A", ""),
        ("Failed Changes", '=COUNTIF(Change_Request_Register!J3:J102,"Failed")', "<5%", '=IF(B7=0,"N/A",IF(B7/B5*100<5,"✓ Within Target","✗ Exceeds Target"))'),
        ("Rolled Back Changes", '=COUNTIF(Change_Request_Register!J3:J102,"Rolled Back")', "<3%", '=IF(B8=0,"N/A",IF(B8/B5*100<3,"✓ Within Target","✗ Exceeds Target"))'),
        ("Changes in Progress", '=COUNTIF(Change_Request_Register!J3:J102,"Implementing")+COUNTIF(Change_Request_Register!J3:J102,"Scheduled")', "N/A", ""),
        ("Overall Success Rate %", '=IF((B6+B7+B8)=0,0,B6/(B6+B7+B8)*100)', "≥95%", '=IF(B10>=95,"✓ Compliant",IF(B10>=90,"\u26A0 Partial","✗ Non-Compliant"))'),
    ]
    
    row += 1
    for metric_name, value_formula, target, status_formula in metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = value_formula
        if "%" in metric_name:
            ws[f'B{row}'].number_format = '0.0'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(name='Calibri', size=11)
        ws[f'C{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        if status_formula:
            ws[f'D{row}'] = status_formula
            apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # SECTION B: Change Type Distribution
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "CHANGE TYPE DISTRIBUTION"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Change Type"
    ws[f'B{row}'] = "Count"
    ws[f'C{row}'] = "Percentage"
    ws[f'D{row}'] = "Target %"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    change_types = [
        ("Standard", '=COUNTIF(Change_Request_Register!C3:C102,"Standard")', "40-50%"),
        ("Normal", '=COUNTIF(Change_Request_Register!C3:C102,"Normal")', "40-50%"),
        ("Emergency", '=COUNTIF(Change_Request_Register!C3:C102,"Emergency")', "<8%"),
        ("Hot Fix", '=COUNTIF(Change_Request_Register!C3:C102,"Hot Fix")', "<2%"),
    ]
    
    row += 1
    total_row_ref = row - 6  # Reference to total changes row (B5)
    for change_type, count_formula, target in change_types:
        ws[f'A{row}'] = change_type
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = count_formula
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = f'=IF($B$5=0,0,B{row}/$B$5*100)'
        ws[f'C{row}'].number_format = '0.0'
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = target
        ws[f'D{row}'].font = Font(name='Calibri', size=11)
        
        row += 1
    
    # SECTION C: Success Rate by Change Type
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "SUCCESS RATE BY CHANGE TYPE"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Change Type"
    ws[f'B{row}'] = "Total"
    ws[f'C{row}'] = "Successful"
    ws[f'D{row}'] = "Success Rate %"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    row += 1
    for change_type, _, _ in change_types:
        ws[f'A{row}'] = change_type
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = f'=COUNTIF(Change_Request_Register!C3:C102,"{change_type}")'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = f'=COUNTIFS(Change_Request_Register!C3:C102,"{change_type}",Change_Request_Register!J3:J102,"Completed")'
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = f'=IF(B{row}=0,0,C{row}/B{row}*100)'
        ws[f'D{row}'].number_format = '0.0'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # SECTION D: Emergency Change Analysis
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "EMERGENCY CHANGE ANALYSIS"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    emerg_metrics = [
        ("Total Emergency Changes", '=COUNTA(Emergency_Changes!A3:A52)-COUNTBLANK(Emergency_Changes!A3:A52)', "N/A", ""),
        ("Emergency Changes %", '=IF($B$5=0,0,B{row}/$B$5*100)', "<10%", '=IF(B{row}<10,"✓ Within Target",IF(B{row}<15,"\u26A0 Warning","✗ Critical"))'),
        ("Emergency Changes Not Reviewed", '=COUNTIF(Emergency_Changes!M3:M52,"Not Yet Reviewed")', "0", '=IF(B{row}=0,"✓ All Reviewed","✗ "&B{row}&" Pending")'),
        ("Process Abuse Flagged", '=COUNTIF(Emergency_Changes!O3:O52,"Likely Abuse")', "0", '=IF(B{row}=0,"✓ None","✗ "&B{row}&" Flagged")'),
    ]
    
    row += 1
    for metric_name, value_formula, target, status_formula in emerg_metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        value_formula = value_formula.replace("{row}", str(row))
        ws[f'B{row}'] = value_formula
        if "%" in metric_name:
            ws[f'B{row}'].number_format = '0.0'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(name='Calibri', size=11)
        ws[f'C{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        status_formula = status_formula.replace("{row}", str(row))
        ws[f'D{row}'] = status_formula
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # Add note
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "NOTE: All metrics auto-calculated. Review monthly. Target: ≥95% success rate, <10% emergency changes."
    ws[f'A{row}'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')

    return ws

def create_compliance_dashboard_sheet(wb, styles):
    """Create Sheet 10: Compliance_Dashboard"""
    ws = wb.create_sheet("Compliance_Dashboard")
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "Change Control Compliance Dashboard"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    # SECTION A: Approval Compliance
    row = 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "APPROVAL COMPLIANCE"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Compliance Check"
    ws[f'B{row}'] = "Compliant"
    ws[f'C{row}'] = "Non-Compliant"
    ws[f'D{row}'] = "Compliance %"
    ws[f'E{row}'] = "Target"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    approval_checks = [
        ("All Changes Have Approval Records", 
         '=COUNTIF(Change_Approval_Workflow!O3:O102,"Approved")+COUNTIF(Change_Approval_Workflow!O3:O102,"Rejected")',
         '=COUNTA(Change_Request_Register!A3:A102)-COUNTBLANK(Change_Request_Register!A3:A102)-B{row}',
         "100%"),
        ("Emergency Changes Have Post-Facto Review",
         '=COUNTIFS(Emergency_Changes!M3:M52,"Approved")+COUNTIFS(Emergency_Changes!M3:M52,"Approved with Remediation")',
         '=COUNTA(Emergency_Changes!A3:A52)-COUNTBLANK(Emergency_Changes!A3:A52)-B{row}',
         "100%"),
    ]
    
    row += 1
    for check_name, compliant_formula, non_compliant_formula, target in approval_checks:
        ws[f'A{row}'] = check_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = compliant_formula
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        non_compliant_formula = non_compliant_formula.replace("{row}", str(row))
        ws[f'C{row}'] = non_compliant_formula
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = f'=IF((B{row}+C{row})=0,0,B{row}/(B{row}+C{row})*100)'
        ws[f'D{row}'].number_format = '0.0'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        ws[f'E{row}'] = target
        ws[f'E{row}'].font = Font(name='Calibri', size=11)
        
        row += 1
    
    # SECTION B: Testing Compliance
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "TESTING COMPLIANCE"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Compliance Check"
    ws[f'B{row}'] = "Compliant"
    ws[f'C{row}'] = "Non-Compliant"
    ws[f'D{row}'] = "Compliance %"
    ws[f'E{row}'] = "Target"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    testing_checks = [
        ("Normal Changes Have Testing",
         '=COUNTIFS(Change_Request_Register!C3:C102,"Normal",Testing_Validation!C3:C102,"Yes")',
         '=COUNTIF(Change_Request_Register!C3:C102,"Normal")-B{row}',
         "100%"),
        ("Test Pass Rate ≥95%",
         '=COUNTIF(Testing_Validation!L3:L102,">=95")',
         '=COUNTIF(Testing_Validation!L3:L102,"<95")',
         "95%"),
        ("Go/No-Go Decision Documented",
         '=COUNTIF(Testing_Validation!P3:P102,"Go")+COUNTIF(Testing_Validation!P3:P102,"No-Go")',
         '=COUNTA(Testing_Validation!A3:A102)-COUNTBLANK(Testing_Validation!A3:A102)-B{row}',
         "100%"),
    ]
    
    row += 1
    for check_name, compliant_formula, non_compliant_formula, target in testing_checks:
        ws[f'A{row}'] = check_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = compliant_formula
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        non_compliant_formula = non_compliant_formula.replace("{row}", str(row))
        ws[f'C{row}'] = non_compliant_formula
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = f'=IF((B{row}+C{row})=0,0,B{row}/(B{row}+C{row})*100)'
        ws[f'D{row}'].number_format = '0.0'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        ws[f'E{row}'] = target
        ws[f'E{row}'].font = Font(name='Calibri', size=11)
        
        row += 1
    
    # SECTION C: Rollback Compliance
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "ROLLBACK COMPLIANCE"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Compliance Check"
    ws[f'B{row}'] = "Compliant"
    ws[f'C{row}'] = "Non-Compliant"
    ws[f'D{row}'] = "Compliance %"
    ws[f'E{row}'] = "Target"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    rollback_checks = [
        ("High-Risk Changes Have Rollback Plan",
         '=COUNTIFS(Impact_Assessment!H3:H102,"High",Rollback_Capability!C3:C102,"Yes")+COUNTIFS(Impact_Assessment!H3:H102,"Critical",Rollback_Capability!C3:C102,"Yes")',
         '=(COUNTIF(Impact_Assessment!H3:H102,"High")+COUNTIF(Impact_Assessment!H3:H102,"Critical"))-B{row}',
         "100%"),
        ("Rollback Procedures Tested",
         '=COUNTIF(Rollback_Capability!P3:P102,"Ready")',
         '=COUNTIF(Rollback_Capability!P3:P102,"Not Ready")',
         "100%"),
    ]
    
    row += 1
    for check_name, compliant_formula, non_compliant_formula, target in rollback_checks:
        ws[f'A{row}'] = check_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = compliant_formula
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        non_compliant_formula = non_compliant_formula.replace("{row}", str(row))
        ws[f'C{row}'] = non_compliant_formula
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = f'=IF((B{row}+C{row})=0,0,B{row}/(B{row}+C{row})*100)'
        ws[f'D{row}'].number_format = '0.0'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        ws[f'E{row}'] = target
        ws[f'E{row}'].font = Font(name='Calibri', size=11)
        
        row += 1
    
    # SECTION D: Overall Summary
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "OVERALL COMPLIANCE SUMMARY"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Category"
    ws[f'D{row}'] = "Compliance %"
    ws[f'E{row}'] = "Status"
    for col in ['A', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    # Calculate averages from sections above
    approval_start = 6
    testing_start = 11
    rollback_start = 17
    
    row += 1
    ws[f'A{row}'] = "Approval Process"
    ws[f'D{row}'] = f'=AVERAGE(D{approval_start}:D{approval_start+1})'
    ws[f'D{row}'].number_format = '0.0'
    apply_style(ws[f'D{row}'], styles['protected_cell'])
    ws[f'E{row}'] = f'=IF(D{row}>=95,"✓ Compliant",IF(D{row}>=90,"\u26A0 Partial","✗ Non-Compliant"))'
    apply_style(ws[f'E{row}'], styles['protected_cell'])
    
    row += 1
    ws[f'A{row}'] = "Testing Process"
    ws[f'D{row}'] = f'=AVERAGE(D{testing_start}:D{testing_start+2})'
    ws[f'D{row}'].number_format = '0.0'
    apply_style(ws[f'D{row}'], styles['protected_cell'])
    ws[f'E{row}'] = f'=IF(D{row}>=95,"✓ Compliant",IF(D{row}>=90,"\u26A0 Partial","✗ Non-Compliant"))'
    apply_style(ws[f'E{row}'], styles['protected_cell'])
    
    row += 1
    ws[f'A{row}'] = "Rollback Readiness"
    ws[f'D{row}'] = f'=AVERAGE(D{rollback_start}:D{rollback_start+1})'
    ws[f'D{row}'].number_format = '0.0'
    apply_style(ws[f'D{row}'], styles['protected_cell'])
    ws[f'E{row}'] = f'=IF(D{row}>=95,"✓ Compliant",IF(D{row}>=90,"\u26A0 Partial","✗ Non-Compliant"))'
    apply_style(ws[f'E{row}'], styles['protected_cell'])
    
    row += 1
    ws[f'A{row}'] = "OVERALL COMPLIANCE"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
    overall_row = row
    ws[f'D{row}'] = f'=AVERAGE(D{overall_row-3}:D{overall_row-1})'
    ws[f'D{row}'].number_format = '0.0'
    ws[f'D{row}'].font = Font(name='Calibri', size=12, bold=True)
    apply_style(ws[f'D{row}'], styles['protected_cell'])
    ws[f'E{row}'] = f'=IF(D{row}>=95,"✓ COMPLIANT",IF(D{row}>=90,"\u26A0 PARTIAL","✗ NON-COMPLIANT"))'
    ws[f'E{row}'].font = Font(name='Calibri', size=12, bold=True)
    apply_style(ws[f'E{row}'], styles['protected_cell'])
    
    # Conditional formatting for compliance %
    ws.conditional_formatting.add(f'D6:D{overall_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=['95'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'D6:D{overall_row}',
        CellIsRule(operator='between', formula=['90', '94.9'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'D6:D{overall_row}',
        CellIsRule(operator='lessThan', formula=['90'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))

    return ws

def create_evidence_register_sheet(wb, styles):
    """Create Sheet 11: Evidence_Register"""
    ws = wb.create_sheet("Evidence_Register")
    
    widths = {
        'A': 18, 'B': 20, 'C': 40, 'D': 20, 'E': 15, 'F': 40,
        'G': 20, 'H': 15, 'I': 15, 'J': 15, 'K': 18, 'L': 25, 'M': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:M1')
    ws['A1'] = "Evidence Register - Change Control Assessment"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    headers = {
        'A': 'Evidence ID', 'B': 'Evidence Type', 'C': 'Evidence Description',
        'D': 'Related Change ID(s)', 'E': 'Evidence Date', 'F': 'Evidence Location',
        'G': 'Evidence Owner', 'H': 'Evidence Classification', 'I': 'Retention Period',
        'J': 'Last Verified Date', 'K': 'Verification Status', 'L': 'Linked Control Requirement',
        'M': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    num_rows = 100
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            apply_style(ws[f'{col}{row}'], styles['data_cell'])
    
    # Data validations
    evid_type_dv = create_data_validation(EVIDENCE_TYPE, allow_blank=False)
    ws.add_data_validation(evid_type_dv)
    evid_type_dv.add(f'B3:B{2+num_rows}')
    
    classif_dv = create_data_validation(EVIDENCE_CLASSIFICATION, allow_blank=False)
    ws.add_data_validation(classif_dv)
    classif_dv.add(f'H3:H{2+num_rows}')
    
    retention_dv = create_data_validation(RETENTION_PERIOD, allow_blank=False)
    ws.add_data_validation(retention_dv)
    retention_dv.add(f'I3:I{2+num_rows}')
    
    verif_dv = create_data_validation(EVIDENCE_VERIFICATION, allow_blank=False)
    ws.add_data_validation(verif_dv)
    verif_dv.add(f'K3:K{2+num_rows}')
    
    # Conditional formatting
    ws.conditional_formatting.add(f'K3:K{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Verified"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'K3:K{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Missing"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    
    ws.conditional_formatting.add(f'H3:H{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Restricted"'], 
                   font=Font(color='9C0006')))
    
    ws.freeze_panes = 'B3'
    
    return ws

def create_approval_signoff_sheet(wb, styles):
    """Create Sheet 12: Approval_Sign_Off"""
    ws = wb.create_sheet("Approval_Sign_Off")
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 60
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "Assessment Approval Sign-Off"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    row = 3
    
    # SECTION A: Document Information
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "DOCUMENT INFORMATION"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    doc_info = [
        ("Assessment Title", "Change Control Assessment - Control A.8.9"),
        ("Assessment Period", "[Start Date DD.MM.YYYY] to [End Date DD.MM.YYYY]"),
        ("Document ID", DOCUMENT_ID),
        ("Version", WORKBOOK_VERSION),
        ("Assessment Date", datetime.now().strftime('%d.%m.%Y')),
    ]
    
    row += 1
    for field, value in doc_info:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = Font(name='Calibri', size=11)
        row += 1
    
    # SECTION B: Preparer Sign-Off
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "PREPARER SIGN-OFF"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    preparer_fields = [
        ("Preparer Name", ""),
        ("Preparer Role", ""),
        ("Preparer Signature", ""),
        ("Date Prepared", ""),
        ("Completeness Attestation", 
         "I attest that all configuration changes have been documented accurately and change control procedures have been followed."),
    ]
    
    row += 1
    for field, default_value in preparer_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws[f'B{row}'] = default_value
        if "Attestation" in field:
            ws[f'B{row}'].font = Font(name='Calibri', size=10, italic=True)
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['info_bg'], end_color=COLORS['info_bg'], fill_type='solid')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 40
        else:
            ws[f'B{row}'].font = Font(name='Calibri', size=11)
        row += 1
    
    # SECTION C: Reviewer Sign-Off
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "REVIEWER SIGN-OFF"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    reviewer_fields = [
        ("Reviewer Name", ""),
        ("Reviewer Role", ""),
        ("Reviewer Signature", ""),
        ("Date Reviewed", ""),
        ("Review Findings", ""),
        ("Gaps Identified", ""),
        ("Review Attestation", 
         "I have reviewed this assessment and verified change control process compliance. Process improvement opportunities have been identified."),
    ]
    
    row += 1
    for field, default_value in reviewer_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws[f'B{row}'] = default_value
        if "Attestation" in field:
            ws[f'B{row}'].font = Font(name='Calibri', size=10, italic=True)
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['info_bg'], end_color=COLORS['info_bg'], fill_type='solid')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 40
        else:
            ws[f'B{row}'].font = Font(name='Calibri', size=11)
            if field in ["Review Findings", "Gaps Identified"]:
                ws.row_dimensions[row].height = 30
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        row += 1
    
    # SECTION D: Approver Sign-Off
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "APPROVER SIGN-OFF"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    approver_fields = [
        ("Approver Name", ""),
        ("Approver Role", ""),
        ("Approver Signature", ""),
        ("Date Approved", ""),
        ("Approval Decision", ""),
        ("Conditions/Comments", ""),
        ("Next Assessment Due", ""),
        ("Approver Attestation", 
         "I approve this change control assessment and authorize remediation activities for identified gaps."),
    ]
    
    row += 1
    decision_row = row + 4  # Track which row has Approval Decision
    for field, default_value in approver_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws[f'B{row}'] = default_value
        if "Attestation" in field:
            ws[f'B{row}'].font = Font(name='Calibri', size=10, italic=True)
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['info_bg'], end_color=COLORS['info_bg'], fill_type='solid')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 40
        else:
            ws[f'B{row}'].font = Font(name='Calibri', size=11)
            if field == "Conditions/Comments":
                ws.row_dimensions[row].height = 30
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        if field == "Approval Decision":
            approval_dv = create_data_validation(APPROVAL_DECISION, allow_blank=True)
            ws.add_data_validation(approval_dv)
            approval_dv.add(f'B{row}')
        
        row += 1
    
    # Conditional formatting for Approval Decision
    ws.conditional_formatting.add(f'B{decision_row}:B{decision_row}',
        CellIsRule(operator='equal', formula=['"Approved"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'B{decision_row}:B{decision_row}',
        CellIsRule(operator='equal', formula=['"Approved with Conditions"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'B{decision_row}:B{decision_row}',
        CellIsRule(operator='equal', formula=['"Not Approved - Revisions Required"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    return ws

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main function to generate the change control assessment workbook."""
    print("=" * 70)
    print(f"Generating {WORKBOOK_TITLE} Workbook")
    print("=" * 70)
    print(f"Document ID: {DOCUMENT_ID}")
    print(f"Version: {WORKBOOK_VERSION}")
    print(f"Date: {datetime.now().strftime('%d.%m.%Y')}")
    print("-" * 70)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create styles
    styles = create_styles()
    
    # Create all sheets
    print("Creating sheets...")
    
    print("  1/12 Creating Instructions sheet...")
    create_instructions_sheet(wb, styles)
    
    print("  2/12 Creating Change_Request_Register sheet (100 rows)...")
    create_change_request_register_sheet(wb, styles)
    
    print("  3/12 Creating Change_Approval_Workflow sheet (100 rows)...")
    create_change_approval_workflow_sheet(wb, styles)
    
    print("  4/12 Creating Impact_Assessment sheet (100 rows)...")
    create_impact_assessment_sheet(wb, styles)
    
    print("  5/12 Creating Testing_Validation sheet (100 rows)...")
    create_testing_validation_sheet(wb, styles)
    
    print("  6/12 Creating Implementation_Log sheet (100 rows)...")
    create_implementation_log_sheet(wb, styles)
    
    print("  7/12 Creating Rollback_Capability sheet (100 rows)...")
    create_rollback_capability_sheet(wb, styles)
    
    print("  8/12 Creating Emergency_Changes sheet (50 rows)...")
    create_emergency_changes_sheet(wb, styles)
    
    print("  9/12 Creating Change_Success_Metrics sheet (dashboard)...")
    create_change_success_metrics_sheet(wb, styles)
    
    print(" 10/12 Creating Compliance_Dashboard sheet...")
    create_compliance_dashboard_sheet(wb, styles)
    
    print(" 11/12 Creating Evidence_Register sheet (100 rows)...")
    create_evidence_register_sheet(wb, styles)
    
    print(" 12/12 Creating Approval_Sign_Off sheet...")
    create_approval_signoff_sheet(wb, styles)
    
    print("  ✓ All sheets created successfully")
    
    # Set workbook properties
    wb.properties.title = WORKBOOK_TITLE
    wb.properties.subject = f"ISMS Control A.8.9 - {WORKBOOK_TITLE}"
    wb.properties.creator = "[Organization] ISMS Implementation Team"
    wb.properties.created = datetime.now()
    wb.properties.description = "Assessment workbook for ISO 27001:2022 Control A.8.9 Change Control requirements"
    
    # Save workbook
    print("-" * 70)
    print("Saving workbook...")
    wb.save(FILENAME)
    
    print("=" * 70)
    print("✓ Workbook generated successfully!")
    print("=" * 70)
    print(f"Output File: {FILENAME}")
    print(f"File Size: {os.path.getsize(FILENAME) / 1024:.1f} KB")
    print(f"Total Sheets: 12")
    print("-" * 70)
    print("\nWorkbook Structure:")
    print("  1.  Instructions - Usage guidance, change types, approval tiers")
    print("  2.  Change_Request_Register - 100 rows for change tracking")
    print("  3.  Change_Approval_Workflow - 100 rows for approval chain")
    print("  4.  Impact_Assessment - 100 rows for risk analysis")
    print("  5.  Testing_Validation - 100 rows for testing records")
    print("  6.  Implementation_Log - 100 rows for execution tracking")
    print("  7.  Rollback_Capability - 100 rows for rollback assessment")
    print("  8.  Emergency_Changes - 50 rows for expedited changes")
    print("  9.  Change_Success_Metrics - Auto-calculated dashboard")
    print("  10. Compliance_Dashboard - Process adherence metrics")
    print("  11. Evidence_Register - 100 rows for evidence documentation")
    print("  12. Approval_Sign_Off - Three-tier approval signatures")
    print("-" * 70)
    print("\nNext Steps:")
    print("1. Open workbook in Excel/LibreOffice")
    print("2. Verify all sheets, validations, and formulas")
    print("3. Review Instructions sheet for change types and priorities")
    print("4. Customize dropdown values if needed (see CONFIGURATION section)")
    print("5. Integrate with existing change management process")
    print("6. Train Change Coordinators on workbook usage")
    print("7. Use dashboards for monthly CAB reporting")
    print("-" * 70)
    print("\nKEY METRICS TO MONITOR:")
    print("\u2022 Change Success Rate: Target ≥95%")
    print("\u2022 Emergency Change Ratio: Target <10%")
    print("\u2022 Approval Compliance: Target 100%")
    print("\u2022 Testing Coverage: Target 100% for Normal changes")
    print("\u2022 Rollback Readiness: Target 100% for High/Critical risk")
    print("\u2022 Overall Compliance: Target ≥95%")
    print("-" * 70)
    print("\nIMPORTANT REMINDERS:")
    print("\u2022 This is a SAMPLE workbook - customize for your organization")
    print("\u2022 Emergency changes require CAB review within 5 business days")
    print("\u2022 High-risk changes require documented rollback procedures")
    print("\u2022 Protected cells (gray) contain formulas - do not edit")
    print("\u2022 Update Change_Request_Register in real-time as changes occur")
    print("\u2022 Review metrics monthly, comprehensive assessment quarterly")
    print("=" * 70)

if __name__ == "__main__":
    main()