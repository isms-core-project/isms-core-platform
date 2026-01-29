#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.9.3 - Configuration Monitoring Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.9: Configuration Management
Assessment Domain 3 of 4: Configuration Monitoring and Drift Detection

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific monitoring infrastructure, drift detection tools,
and remediation workflows.

Key customization areas:
1. Monitoring tool capabilities (match your actual deployment)
2. Drift detection thresholds (adapt to your risk tolerance)
3. Alert routing and escalation (align with your SOC/operations structure)
4. Remediation SLAs (based on your operational requirements)
5. Monitoring coverage targets (specific to your asset criticality tiers)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.9 Configuration Management Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
configuration monitoring and drift detection capabilities against ISO 27001:2022
Control A.8.9 requirements.

**Purpose:**
Enables systematic assessment of continuous monitoring, drift detection
effectiveness, alert management, and remediation processes to ensure
unauthorized configuration changes are detected and corrected.

**Assessment Scope:**
- Monitoring infrastructure deployment and coverage
- Drift detection capabilities and thresholds
- Alert generation and routing effectiveness
- Remediation workflow execution and SLA compliance
- Baseline comparison and compliance verification
- Monitoring tool performance and reliability
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and monitoring standards
2. Monitoring Infrastructure - Tool deployment and coverage tracking
3. Drift Detection - Alert configuration and detection capabilities
4. Baseline Comparison - Compliance scan results and drift analysis
5. Remediation Tracking - Drift incident remediation and SLA compliance
6. Monitoring Performance - Tool uptime, alert accuracy, false positive rates
7. Gap Analysis - Coverage gaps and remediation requirements
8. Evidence Register - Audit evidence tracking (100+ entries)
9. Approval & Sign-Off - Three-tier approval workflow

**Key Features:**
- Data validation with drift severity and asset tier dropdowns
- Conditional formatting for coverage status and SLA compliance
- Automated gap identification for unmonitored assets
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with monitoring and SIEM systems

**Integration:**
This assessment feeds into the A.8.9.5 Compliance Dashboard, which
consolidates data from all four configuration management assessment domains
for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a89_3_monitoring.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a89_3_monitoring.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a89_3_monitoring.py --date 20250127

Output:
    File: ISMS_IMP_A_8_9_3_Configuration_Monitoring_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize monitoring coverage targets by asset tier
    2. Inventory all deployed monitoring tools and agents
    3. Document drift detection capabilities and thresholds
    4. Assess baseline comparison scan frequency and coverage
    5. Verify alert routing and remediation workflows
    6. Calculate remediation SLA compliance rates
    7. Review monitoring tool performance metrics
    8. Conduct gap analysis for unmonitored assets
    9. Define remediation actions with timelines
    10. Collect and link audit evidence (drift alerts, scan results)
    11. Obtain three-tier stakeholder approvals
    12. Feed results into A.8.9.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.9
Assessment Domain:    3 of 4 (Configuration Monitoring and Drift Detection)
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.9: Configuration Management Policy (Consolidated)
    - ISMS-POL-A.8.9, Section 2.4: Configuration Monitoring & Drift Detection
    - ISMS-CTX-A.8.9: Configuration Management Reference (NOT ISMS)
    - ISMS-CTX-A.8.9, Part 3: Configuration Deviation Response Procedures
    - ISMS-CTX-A.8.9-Evidence-Collection: Evidence Collection Guide
    - ISMS-IMP-A.8.9.1: Baseline Configuration Assessment (Domain 1)
    - ISMS-IMP-A.8.9.2: Change Control Assessment (Domain 2)
    - ISMS-IMP-A.8.9.4: Security Hardening Assessment (Domain 4)
    - ISMS-IMP-A.8.9.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.9.3 specification
    - Supports comprehensive configuration monitoring evaluation
    - Integrated with A.8.9.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Monitoring Technology:**
Configuration monitoring tools and techniques evolve rapidly. Review vendor
roadmaps, emerging monitoring capabilities, and drift detection algorithms
quarterly. False positive tuning and alert fatigue prevention are critical.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of monitoring coverage, drift detection, and
remediation effectiveness.

**Data Protection:**
Assessment workbooks contain sensitive operational details including:
- Monitoring tool deployment architecture
- Drift alert details with system configurations
- Baseline compliance scan results
- Vulnerability information and security gaps

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Monthly: Check monitoring coverage and drift remediation SLA compliance
- Quarterly: Review drift trends and monitoring tool performance
- Annually: Complete reassessment of monitoring infrastructure
- Ad-hoc: When monitoring tools change or new asset types deployed

**Quality Assurance:**
Have SOC analysts, monitoring engineers, and configuration managers validate
assessments before using results for compliance reporting or tool investment.

**Regulatory Alignment:**
Ensure monitoring practices align with applicable requirements:
- ISO 27001:2022: Control A.8.9 monitoring requirements
- Continuous monitoring: Real-time security monitoring mandates
- Sector-specific: Regulatory monitoring and alerting requirements
- Internal: Organizational monitoring and incident response policies

Customize assessment criteria to include regulatory-specific requirements.

================================================================================
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •

from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

# ============================================================================
# CONFIGURATION SECTION - CUSTOMIZE FOR YOUR ENVIRONMENT
# ============================================================================

# File output configuration
FILENAME = f"ISMS-IMP-A.8.X.X_..._Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"

# Workbook metadata
WORKBOOK_TITLE = "Configuration Monitoring Assessment"
WORKBOOK_VERSION = "1.0"
DOCUMENT_ID = "ISMS-IMP-A.8.9.3"

# CUSTOMIZE: Configuration monitoring dropdown values
ASSET_CRITICALITY = ["🔴 Critical", "🟡 High", "🟢 Medium", "⭕ Low"]
MONITORING_STATUS = ["\u2705 Monitored", "\u26A0\uFE0F Partially Monitored", "\u274C Not Monitored", "➖ Excluded"]
MONITORING_METHOD = ["Automated Continuous", "Scheduled Automated", "Manual", "Hybrid", "None"]
CHECK_FREQUENCY = ["Real-time (<15 min)", "Hourly", "Daily", "Weekly", "Monthly", "Quarterly", "Manual (on-demand)"]

DRIFT_CATEGORY = ["🔴 Critical", "🟡 High", "🟢 Medium", "⭕ Low", "ℹ️ Informational"]
DETECTION_METHOD = ["Automated Continuous", "Scheduled Scan", "Manual Check", "User Report"]
AUTHORIZED_CHANGE = ["Yes (Change ID)", "No (Unauthorized)", "Under Investigation"]
ROOT_CAUSE_CATEGORY = ["Unauthorized Manual Change", "Tool Failure", "Software Update", 
                       "Baseline Not Updated", "Environmental", "Malicious", "Other"]
DRIFT_STATUS = ["🔍 Detected", "🔎 Under Investigation", "⏳ Remediation In Progress", 
               "\u2705 Remediated", "✔️ Closed", "➖ False Positive"]

TOOL_TYPE = ["Agent-Based", "Agentless", "Network Scanner", "Script/Custom", 
            "Cloud-Native", "SIEM Integration"]
DEPLOYMENT_STATUS = ["\u2705 Active", "\u26A0\uFE0F Degraded", "\u274C Offline", "🧪 Pilot", "⏸️ Decommissioned"]
ALERTING_METHOD = ["Email", "SIEM", "Webhook", "Dashboard Only", "Ticketing System", "Multiple"]
LICENSING_MODEL = ["Commercial", "Open Source", "Subscription", "Perpetual", "In-House Developed"]

REMEDIATION_ACTION = ["Reverted to Baseline", "Updated Baseline", "Authorized Retroactively", 
                     "No Action (False Positive)", "Other"]
VERIFICATION_METHOD = ["Automated Re-Scan", "Manual Verification", "Monitoring Tool Confirmation", 
                      "User Validation"]
VERIFICATION_RESULT = ["\u2705 Passed", "\u274C Failed", "\u26A0\uFE0F Partially Successful", "❓ Not Yet Verified"]
ROOT_CAUSE_REMEDIATION = ["Baseline Updated", "Change Control Enforced", "Tool Fixed", 
                         "Process Improved", "Training Provided", "Other"]

FALSE_POSITIVE_REASON = ["Incorrect Baseline", "Tool Misconfiguration", "Expected Variation", 
                        "Timing Issue", "Tool Bug", "Other"]
TUNING_ACTION = ["Baseline Updated", "Monitoring Rule Adjusted", "Alert Threshold Changed", 
                "Exception Added", "Tool Updated", "No Action (Acceptable)"]
RECURRENCE_STATUS = ["Not Seen Again", "Recurred Once", "Recurring (Needs Further Tuning)", "Monitoring"]

EVIDENCE_TYPE = ["Monitoring Configuration", "Drift Alert Screenshot", "Remediation Record", 
                "Tool Health Check", "Coverage Report", "Scan Output", "Other"]
EVIDENCE_CLASSIFICATION = ["Public", "Internal", "Confidential", "Restricted"]
RETENTION_PERIOD = ["1 Year", "3 Years", "5 Years", "7 Years", "Indefinite"]
EVIDENCE_VERIFICATION = ["Verified", "Needs Verification", "Missing", "Outdated"]

APPROVAL_DECISION = ["Approved", "Approved with Conditions", "Not Approved - Revisions Required"]

# CUSTOMIZE: Asset types (43 types from A.8.9.1)
ASSET_TYPES = {
    "Infrastructure": [
        "Physical Server", "Virtual Machine", "Hypervisor", "Network Switch",
        "Router", "Firewall", "Load Balancer", "Storage Array",
        "Backup Appliance", "UPS", "HVAC", "Physical Security System"
    ],
    "Endpoint": [
        "Desktop", "Laptop", "Mobile Phone", "Tablet", "Thin Client", "Kiosk/POS"
    ],
    "Network Services": [
        "DNS Server", "DHCP Server", "NTP Server", "Proxy Server",
        "VPN Gateway", "Wireless Access Point", "Network Management System"
    ],
    "Applications": [
        "Enterprise Application", "Web Application", "Database Server",
        "Middleware", "API Gateway", "Custom Developed Application",
        "COTS Application", "Open Source Application"
    ],
    "Cloud": [
        "IaaS Resource", "PaaS Service", "SaaS Application",
        "Cloud-Native Application", "Serverless Function"
    ],
    "IoT/OT": [
        "IoT Device", "Industrial Control System", "SCADA System",
        "Building Management System", "Medical Device"
    ]
}

# CUSTOMIZE: Color scheme
COLORS = {
    'header_main': '003366',
    'header_sub': '4472C4',
    'column_header': 'D9D9D9',
    'input_cell': 'FFFFFF',
    'protected_cell': 'F2F2F2',
    'compliant': 'C6EFCE',
    'partial': 'FFEB9C',
    'non_compliant': 'FFC7CE',
    'excluded': 'D9D9D9',
    'critical': 'C00000',
    'info_bg': 'E7E6E6',
    'light_green': 'E2EFDA',
    'orange': 'FFA500'
}

# ============================================================================
# STYLE DEFINITIONS
# ============================================================================

def create_styles():
    """Creates and returns a dictionary of reusable styles."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    styles = {
        'header_main': {
            'font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_main'], 
                              end_color=COLORS['header_main'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'header_sub': {
            'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_sub'], 
                              end_color=COLORS['header_sub'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'column_header': {
            'font': Font(name='Calibri', size=11, bold=True),
            'fill': PatternFill(start_color=COLORS['column_header'], 
                              end_color=COLORS['column_header'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': thin_border
        },
        'data_cell': {
            'font': Font(name='Calibri', size=11),
            'fill': PatternFill(start_color=COLORS['input_cell'], 
                              end_color=COLORS['input_cell'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=False),
            'border': thin_border
        },
        'protected_cell': {
            'font': Font(name='Calibri', size=11, italic=True),
            'fill': PatternFill(start_color=COLORS['protected_cell'], 
                              end_color=COLORS['protected_cell'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top'),
            'border': thin_border
        },
        'info_text': {
            'font': Font(name='Calibri', size=11),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True)
        },
        'section_header': {
            'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color=COLORS['header_sub'], 
                              end_color=COLORS['header_sub'], 
                              fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': thin_border
        }
    }
    return styles

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

def set_column_widths(ws, widths):
    """Set column widths for a worksheet."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def create_data_validation(values, allow_blank=True):
    """Create a data validation object for dropdowns."""
    formula = f'"{",".join(values)}"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = 'Invalid entry'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select from the dropdown'
    dv.promptTitle = 'Selection Required'
    return dv

def protect_formula_cells(ws, start_row, end_row, formula_columns):
    """Protect cells containing formulas."""
    for row in range(start_row, end_row + 1):
        for col in formula_columns:
            ws[f'{col}{row}'].protection = Protection(locked=True)

# ============================================================================
# SHEET CREATION FUNCTIONS
# ============================================================================

def create_lookup_tables(wb, styles):
    """Create hidden Lookup_Tables sheet with 43 asset types."""
    ws = wb.create_sheet("Lookup_Tables")
    ws.sheet_state = 'hidden'
    
    ws['A1'] = "Asset Type"
    ws['B1'] = "Asset Category"
    apply_style(ws['A1'], styles['column_header'])
    apply_style(ws['B1'], styles['column_header'])
    
    row = 2
    for category, types in ASSET_TYPES.items():
        for asset_type in types:
            ws[f'A{row}'] = asset_type
            ws[f'B{row}'] = category
            row += 1
    
    return ws

def create_instructions_sheet(wb, styles):
    """Create the Instructions and Legend sheet."""
    ws = wb.create_sheet("Instructions", 0)
    ws.column_dimensions['A'].width = 100
    
    # Title
    ws.merge_cells('A1:A2')
    ws['A1'] = "ISMS Control A.8.9 - Configuration Monitoring Assessment"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color=COLORS['header_main'], 
                                end_color=COLORS['header_main'], 
                                fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Document metadata
    ws['A3'] = "Document ID:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = DOCUMENT_ID
    
    ws['A4'] = "Assessment:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "Configuration Monitoring Assessment"
    
    ws['A5'] = "Version:"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = "1.0"
    
    ws['A6'] = "Generated:"
    ws['A6'].font = Font(bold=True)
    ws['B6'] = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    ws.column_dimensions['B'].width = 40
    
    # CUSTOMIZE: Update instructions for your organization
    instructions = [
        "",
        "ASSESSMENT OVERVIEW",
        "━" * 80,
        "",
        "Purpose:",
        "This workbook evaluates configuration monitoring and drift detection effectiveness. "
        "It verifies that configuration states are continuously monitored against approved baselines, "
        "deviations are detected promptly, and corrective actions are tracked. Provides evidence of "
        "ISO 27001:2022 Control A.8.9 monitoring compliance.",
        "",
        "Assessment Scope:",
        "\u2022 Configuration monitoring coverage (which assets are monitored)",
        "\u2022 Drift detection capabilities (automated vs. manual)",
        "\u2022 Baseline compliance verification frequency",
        "\u2022 Drift alerting and notification workflows",
        "\u2022 Drift incident tracking and remediation",
        "\u2022 False positive management",
        "\u2022 Monitoring tool effectiveness",
        "\u2022 Integration with change control (authorized vs. unauthorized changes)",
        "",
        "MONITORING METHODS",
        "━" * 80,
        "",
        "Automated Continuous Monitoring:",
        "\u2022 Real-time or near-real-time drift detection",
        "\u2022 Agent-based tools or agentless API polling",
        "\u2022 Examples: Configuration management tools, SIEM integrations, cloud-native monitoring",
        "\u2022 Frequency: Continuous (minutes to hours)",
        "\u2022 Best for: Large-scale, dynamic environments, critical assets",
        "",
        "Scheduled Automated Scans:",
        "\u2022 Periodic automated configuration checks",
        "\u2022 Examples: Daily/weekly compliance scans, scheduled scripts",
        "\u2022 Frequency: Daily, weekly, or monthly",
        "\u2022 Best for: Stable environments with infrequent changes",
        "",
        "Manual Verification:",
        "\u2022 Human review of configuration files or settings",
        "\u2022 Examples: Quarterly configuration reviews, post-change verification",
        "\u2022 Frequency: Weekly, monthly, or quarterly",
        "\u2022 Best for: Critical systems requiring human validation, spot checks",
        "",
        "Hybrid Approach:",
        "\u2022 Combination of automated and manual methods",
        "\u2022 Examples: Automated daily scans + weekly manual review of critical items",
        "\u2022 Best for: Balancing coverage with resource constraints",
        "",
        "DRIFT SEVERITY CATEGORIES",
        "━" * 80,
        "",
        "Critical Drift (Response: <4 hours):",
        "\u2022 Security control disabled",
        "\u2022 Public exposure created (firewall rule allowing unrestricted access)",
        "\u2022 Encryption disabled",
        "\u2022 Immediate investigation and remediation required",
        "",
        "High Drift (Response: <24 hours):",
        "\u2022 Significant security baseline deviation",
        "\u2022 Security patch removed",
        "\u2022 Authentication weakened",
        "\u2022 Unauthorized service enabled",
        "",
        "Medium Drift (Response: <7 days):",
        "\u2022 Moderate deviation with potential security impact",
        "\u2022 Non-security configuration change",
        "\u2022 Logging level reduced",
        "\u2022 Resource limit modified",
        "",
        "Low Drift (Response: <30 days):",
        "\u2022 Minor deviation with minimal security impact",
        "\u2022 Cosmetic settings changed",
        "\u2022 Non-critical parameter drift",
        "\u2022 Timezone discrepancy",
        "",
        "Informational:",
        "\u2022 Benign drift, no action required",
        "\u2022 Expected variation (e.g., dynamic IP)",
        "\u2022 Authorized temporary change documented",
        "",
        "MONITORING COVERAGE TIERS",
        "━" * 80,
        "",
        "Tier 1 - Critical Assets (Continuous Monitoring Required):",
        "\u2022 Public-facing systems (web servers, APIs)",
        "\u2022 Authentication infrastructure (directory services, IAM)",
        "\u2022 Security controls (firewalls, IDS/IPS, SIEM)",
        "\u2022 Database servers with sensitive data",
        "\u2022 Monitoring: Real-time or <15 minute intervals",
        "\u2022 Target: 100% coverage, 100% detection within 1 hour",
        "",
        "Tier 2 - High Value Assets (Frequent Monitoring):",
        "\u2022 Internal application servers",
        "\u2022 File and storage servers",
        "\u2022 Network infrastructure (switches, routers)",
        "\u2022 Monitoring: Hourly to daily",
        "\u2022 Target: ≥95% coverage, 95% detection within 24 hours",
        "",
        "Tier 3 - Standard Assets (Regular Monitoring):",
        "\u2022 Workstations and laptops",
        "\u2022 Non-critical application servers",
        "\u2022 Development/test environments",
        "\u2022 Monitoring: Daily to weekly",
        "\u2022 Target: ≥85% coverage, 90% detection within 7 days",
        "",
        "Tier 4 - Low Risk Assets (Periodic Monitoring):",
        "\u2022 Isolated lab systems",
        "\u2022 Non-connected devices",
        "\u2022 Monitoring: Weekly to monthly",
        "\u2022 Target: ≥70% coverage",
        "",
        "WHO SHOULD COMPLETE THIS ASSESSMENT",
        "━" * 80,
        "",
        "System Administrators / SOC Analysts (Preparer):",
        "\u2022 Document monitoring coverage per asset",
        "\u2022 Record drift detection incidents as they occur",
        "\u2022 Track remediation actions",
        "\u2022 Monitor false positive rates",
        "\u2022 Maintain evidence of monitoring tool operation",
        "",
        "Configuration Manager / IT Operations Manager (Reviewer):",
        "\u2022 Verify monitoring coverage adequacy",
        "\u2022 Analyze drift trends and patterns",
        "\u2022 Assess detection effectiveness (MTTD)",
        "\u2022 Review remediation timeliness",
        "\u2022 Identify monitoring gaps",
        "\u2022 Recommend tool/process improvements",
        "",
        "CISO / IT Manager (Approver):",
        "\u2022 Review overall monitoring effectiveness",
        "\u2022 Approve monitoring coverage expansion plans",
        "\u2022 Authorize budget for monitoring tools",
        "\u2022 Sign off on assessment",
        "",
        "COMPLETION INSTRUCTIONS",
        "━" * 80,
        "",
        "Ongoing Operations (Continuous):",
        "1. Document all assets in Monitoring_Coverage_Register",
        "2. Log drift incidents in Drift_Detection_Log as detected",
        "3. Investigate unauthorized changes immediately",
        "4. Track remediation in Drift_Remediation_Tracking",
        "5. Document false positives in False_Positive_Register",
        "6. Tune monitoring rules to reduce noise",
        "",
        "Monthly Review (Configuration Manager):",
        "1. Review Monitoring_Effectiveness_Metrics",
        "2. Analyze drift trends from previous month",
        "3. Verify SLA compliance for remediation",
        "4. Review false positive rates and implement tuning",
        "5. Update monitoring coverage for new assets",
        "",
        "Quarterly Assessment (Configuration Manager):",
        "1. Comprehensive review of all sheets",
        "2. Analyze Drift_Trend_Analysis for patterns",
        "3. Review Coverage_Gap_Analysis priorities",
        "4. Assess monitoring tool effectiveness",
        "5. Compile evidence and complete Reviewer Sign-Off",
        "",
        "Annual Approval (CISO/IT Manager):",
        "1. Review monitoring program effectiveness",
        "2. Approve tool budget and coverage expansion",
        "3. Complete Approver Sign-Off",
        "",
        "WORKBOOK STRUCTURE",
        "━" * 80,
        "",
        "This workbook contains 11 sheets:",
        "",
        "1.  Instructions - This sheet (usage guidance)",
        "2.  Monitoring_Coverage_Register - Asset monitoring inventory (100 rows)",
        "3.  Drift_Detection_Log - Drift incident records (150 rows)",
        "4.  Monitoring_Tool_Inventory - Tools and capabilities (30 rows)",
        "5.  Drift_Remediation_Tracking - Remediation actions (150 rows)",
        "6.  False_Positive_Register - False alert tracking (75 rows)",
        "7.  Monitoring_Effectiveness_Metrics - Auto-calculated KPIs (dashboard)",
        "8.  Coverage_Gap_Analysis - Coverage analysis (dashboard)",
        "9.  Drift_Trend_Analysis - Temporal trends (dashboard)",
        "10. Evidence_Register - Supporting evidence (100 rows)",
        "11. Approval_Sign_Off - Three-tier signatures",
        "",
        "LEGEND - COLOR CODING",
        "━" * 80,
        "",
        "Drift Severity:",
        "  🔴 Critical (Dark Red) - Immediate action <4 hours",
        "  🔴 High (Red) - Urgent action <24 hours",
        "  🟡 Medium (Yellow) - Important <7 days",
        "  🟢 Low (Light Green) - Minor <30 days",
        "  ⚪ Informational (Gray) - No action required",
        "",
        "Monitoring Status:",
        "  🟢 Monitored (Green) - Active monitoring in place",
        "  🟡 Partially Monitored (Yellow) - Some coverage gaps",
        "  🔴 Not Monitored (Red) - No monitoring",
        "  ⚪ Excluded (Gray) - Out of scope",
        "",
        "Coverage Compliance:",
        "  🟢 Compliant (Green) - Meets tier requirements",
        "  🟡 Partial (Yellow) - Some gaps acceptable",
        "  🔴 Non-Compliant (Red) - Critical gap",
        "",
        "INTEGRATION WITH OTHER ASSESSMENTS",
        "━" * 80,
        "",
        "Links to ISMS-IMP-A.8.9.1 (Baseline Configuration):",
        "\u2022 Asset IDs should match Asset_Inventory",
        "\u2022 Baseline Reference links to Baseline_Repository",
        "\u2022 Expected values come from documented baselines",
        "",
        "Links to ISMS-IMP-A.8.9.2 (Change Control):",
        "\u2022 Authorized changes reference Change IDs",
        "\u2022 Drift incidents verified against Change_Request_Register",
        "\u2022 Unauthorized changes trigger change control review",
        "",
        "COMPLIANCE TARGETS",
        "━" * 80,
        "",
        "Monitoring Coverage:",
        "\u2022 Tier 1 (Critical): 100%",
        "\u2022 Tier 2 (High): ≥95%",
        "\u2022 Tier 3 (Standard): ≥85%",
        "\u2022 Tier 4 (Low): ≥70%",
        "\u2022 Overall: ≥85%",
        "",
        "Detection Performance:",
        "\u2022 Mean Time to Detect (MTTD): <1 hour (Tier 1), <24 hours (Tier 2)",
        "\u2022 False Positive Rate: <10%",
        "",
        "Remediation Performance:",
        "\u2022 Critical Drift SLA: <4 hours (100% compliance)",
        "\u2022 High Drift SLA: <24 hours (≥95% compliance)",
        "\u2022 Overall Success Rate: ≥98%",
        "",
        "IMPORTANT NOTES",
        "━" * 80,
        "",
        "\u2022 Protected cells (gray) contain formulas - do not edit",
        "\u2022 Critical drift (dark red) requires immediate escalation",
        "\u2022 Log all drift incidents even if immediately remediated",
        "\u2022 False positives must be analyzed and tuned",
        "\u2022 Dashboards auto-update - review monthly",
        "\u2022 Tools marked 'Offline' create monitoring gaps",
        "\u2022 Integration with A.8.9.1 and A.8.9.2 is critical for traceability",
        "",
        "SUPPORT AND QUESTIONS",
        "━" * 80,
        "",
        "Configuration Manager: [contact information]",
        "Security Operations Center: [contact information]",
        "ISMS Team: [contact information]",
        "",
        "For technical issues: [IT support]",
        "For policy questions: Reference ISMS-POL-A.8.9-S2.3",
        "",
        "━" * 80,
        f"Generated: {datetime.now().strftime('%d.%m.%Y')} | Version: {WORKBOOK_VERSION} | Document ID: {DOCUMENT_ID}",
    ]
    
    row = 8
    for line in instructions:
        ws[f'A{row}'] = line
        if line.startswith("━"):
            ws[f'A{row}'].font = Font(name='Calibri', size=11, color='666666')
        elif line.isupper() and len(line) > 5 and not line.startswith("  "):
            apply_style(ws[f'A{row}'], styles['section_header'])
            ws.row_dimensions[row].height = 20
        else:
            apply_style(ws[f'A{row}'], styles['info_text'])
        row += 1
    
    return ws

def create_monitoring_coverage_register_sheet(wb, styles):
    """Create Sheet 2: Monitoring_Coverage_Register"""
    ws = wb.create_sheet("Monitoring_Coverage_Register")
    
    widths = {
        'A': 18,  # Asset ID
        'B': 30,  # Asset Name
        'C': 35,  # Asset Type
        'D': 20,  # Asset Category
        'E': 15,  # Asset Criticality
        'F': 25,  # Monitoring Tier
        'G': 18,  # Monitoring Status
        'H': 25,  # Monitoring Method
        'I': 30,  # Monitoring Tool/System
        'J': 20,  # Check Frequency
        'K': 25,  # Baseline Reference
        'L': 15,  # Last Monitored Date
        'M': 35,  # Monitoring Configuration Location
        'N': 18,  # Coverage Compliance
        'O': 40,  # Gap Justification
        'P': 35   # Notes
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:P1')
    ws['A1'] = "Configuration Monitoring Coverage Register"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    headers = {
        'A': 'Asset ID', 'B': 'Asset Name', 'C': 'Asset Type', 'D': 'Asset Category',
        'E': 'Asset Criticality', 'F': 'Monitoring Tier', 'G': 'Monitoring Status',
        'H': 'Monitoring Method', 'I': 'Monitoring Tool/System', 'J': 'Check Frequency',
        'K': 'Baseline Reference', 'L': 'Last Monitored Date', 
        'M': 'Monitoring Configuration Location', 'N': 'Coverage Compliance',
        'O': 'Gap Justification', 'P': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    num_rows = 100
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['D', 'F', 'N']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Formulas
    for row in range(3, 3 + num_rows):
        # Column D: Asset Category (VLOOKUP from Lookup_Tables)
        ws[f'D{row}'] = f'=IFERROR(VLOOKUP(C{row},Lookup_Tables!$A$2:$B$44,2,FALSE),"")'
        
        # Column F: Monitoring Tier (based on criticality)
        ws[f'F{row}'] = f'=IF(E{row}="Critical","Tier 1 - Critical Assets",IF(E{row}="High","Tier 2 - High Value Assets",IF(E{row}="Medium","Tier 3 - Standard Assets","Tier 4 - Low Risk Assets")))'
        
        # Column N: Coverage Compliance (complex logic)
        ws[f'N{row}'] = f'=IF(G{row}="Excluded","Excluded",IF(E{row}="Critical",IF(AND(G{row}="Monitored",OR(H{row}="Automated Continuous",H{row}="Hybrid")),"Compliant","Non-Compliant"),IF(E{row}="High",IF(G{row}="Monitored","Compliant",IF(G{row}="Partially Monitored","Partial","Non-Compliant")),IF(G{row}="Not Monitored","Non-Compliant","Compliant"))))'
    
    # Data validations
    asset_type_dv = DataValidation(type="list", formula1="Lookup_Tables!$A$2:$A$44", allow_blank=True)
    ws.add_data_validation(asset_type_dv)
    asset_type_dv.add(f'C3:C{2+num_rows}')
    
    criticality_dv = create_data_validation(ASSET_CRITICALITY, allow_blank=False)
    ws.add_data_validation(criticality_dv)
    criticality_dv.add(f'E3:E{2+num_rows}')
    
    status_dv = create_data_validation(MONITORING_STATUS, allow_blank=False)
    ws.add_data_validation(status_dv)
    status_dv.add(f'G3:G{2+num_rows}')
    
    method_dv = create_data_validation(MONITORING_METHOD, allow_blank=False)
    ws.add_data_validation(method_dv)
    method_dv.add(f'H3:H{2+num_rows}')
    
    freq_dv = create_data_validation(CHECK_FREQUENCY, allow_blank=True)
    ws.add_data_validation(freq_dv)
    freq_dv.add(f'J3:J{2+num_rows}')
    
    # Conditional formatting
    ws.conditional_formatting.add(f'E3:E{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Critical"'], 
                   fill=PatternFill(start_color=COLORS['critical'], end_color=COLORS['critical'], fill_type='solid'),
                   font=Font(bold=True, color='FFFFFF')))
    ws.conditional_formatting.add(f'E3:E{2+num_rows}',
        CellIsRule(operator='equal', formula=['"High"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'E3:E{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Medium"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'E3:E{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Low"'], 
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'G3:G{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Monitored"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'G3:G{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Not Monitored"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'N3:N{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Compliant"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'N3:N{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Non-Compliant"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    
    ws.freeze_panes = 'B3'
    protect_formula_cells(ws, 3, 2 + num_rows, ['D', 'F', 'N'])
    
    return ws

def create_drift_detection_log_sheet(wb, styles):
    """Create Sheet 3: Drift_Detection_Log"""
    ws = wb.create_sheet("Drift_Detection_Log")
    
    widths = {
        'A': 20, 'B': 20, 'C': 18, 'D': 25, 'E': 30, 'F': 25,
        'G': 25, 'H': 15, 'I': 20, 'J': 25, 'K': 20, 'L': 25,
        'M': 25, 'N': 35, 'O': 20, 'P': 12, 'Q': 20, 'R': 15, 'S': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:S1')
    ws['A1'] = "Configuration Drift Detection Log"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    headers = {
        'A': 'Drift Incident ID', 'B': 'Detection Date/Time', 'C': 'Asset ID',
        'D': 'Asset Name', 'E': 'Configuration Item', 'F': 'Expected Value',
        'G': 'Actual Value', 'H': 'Drift Category', 'I': 'Detection Method',
        'J': 'Detecting Tool/System', 'K': 'Authorized Change', 'L': 'Change ID Reference',
        'M': 'Root Cause Category', 'N': 'Root Cause Detail', 'O': 'Drift Status',
        'P': 'Time to Detect (Hours)', 'Q': 'Remediation Owner', 'R': 'Priority', 'S': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    num_rows = 150
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['P', 'R']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Formulas
    for row in range(3, 3 + num_rows):
        # Column P: Time to Detect (manual in this simplified version)
        ws[f'P{row}'] = f'=IF(B{row}="","","")'
        
        # Column R: Priority
        ws[f'R{row}'] = f'=IF(H{row}="Critical","P1-Critical",IF(H{row}="High","P2-High",IF(H{row}="Medium","P3-Medium","P4-Low")))'
    
    # Data validations
    drift_cat_dv = create_data_validation(DRIFT_CATEGORY, allow_blank=False)
    ws.add_data_validation(drift_cat_dv)
    drift_cat_dv.add(f'H3:H{2+num_rows}')
    
    detect_dv = create_data_validation(DETECTION_METHOD, allow_blank=False)
    ws.add_data_validation(detect_dv)
    detect_dv.add(f'I3:I{2+num_rows}')
    
    auth_dv = create_data_validation(AUTHORIZED_CHANGE, allow_blank=False)
    ws.add_data_validation(auth_dv)
    auth_dv.add(f'K3:K{2+num_rows}')
    
    root_dv = create_data_validation(ROOT_CAUSE_CATEGORY, allow_blank=True)
    ws.add_data_validation(root_dv)
    root_dv.add(f'M3:M{2+num_rows}')
    
    status_dv = create_data_validation(DRIFT_STATUS, allow_blank=False)
    ws.add_data_validation(status_dv)
    status_dv.add(f'O3:O{2+num_rows}')
    
    # Conditional formatting
    ws.conditional_formatting.add(f'H3:H{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Critical"'], 
                   fill=PatternFill(start_color=COLORS['critical'], end_color=COLORS['critical'], fill_type='solid'),
                   font=Font(bold=True, color='FFFFFF')))
    ws.conditional_formatting.add(f'H3:H{2+num_rows}',
        CellIsRule(operator='equal', formula=['"High"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'H3:H{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Medium"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'K3:K{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Yes (Change ID)"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'K3:K{2+num_rows}',
        CellIsRule(operator='equal', formula=['"No (Unauthorized)"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'O3:O{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Closed"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    
    ws.freeze_panes = 'B3'
    protect_formula_cells(ws, 3, 2 + num_rows, ['P', 'R'])
    
    return ws

def create_monitoring_tool_inventory_sheet(wb, styles):
    """Create Sheet 4: Monitoring_Tool_Inventory"""
    ws = wb.create_sheet("Monitoring_Tool_Inventory")
    
    widths = {
        'A': 15,  # Tool ID
        'B': 25,  # Tool Name
        'C': 20,  # Vendor
        'D': 20,  # Tool Type
        'E': 40,  # Monitoring Capabilities
        'F': 35,  # Asset Types Covered
        'G': 15,  # Assets Monitored Count
        'H': 18,  # Deployment Status
        'I': 35,  # Integration Points
        'J': 20,  # Alerting Method
        'K': 20,  # Licensing Model
        'L': 15,  # Annual Cost
        'M': 15,  # Last Health Check
        'N': 35,  # Known Limitations
        'O': 20,  # Tool Owner
        'P': 35   # Notes
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:P1')
    ws['A1'] = "Monitoring Tool Inventory"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    headers = {
        'A': 'Tool ID', 'B': 'Tool Name', 'C': 'Vendor', 'D': 'Tool Type',
        'E': 'Monitoring Capabilities', 'F': 'Asset Types Covered',
        'G': 'Assets Monitored Count', 'H': 'Deployment Status',
        'I': 'Integration Points', 'J': 'Alerting Method', 'K': 'Licensing Model',
        'L': 'Annual Cost', 'M': 'Last Health Check', 'N': 'Known Limitations',
        'O': 'Tool Owner', 'P': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    num_rows = 30
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            apply_style(ws[f'{col}{row}'], styles['data_cell'])
    
    # Data validations
    tool_type_dv = create_data_validation(TOOL_TYPE, allow_blank=False)
    ws.add_data_validation(tool_type_dv)
    tool_type_dv.add(f'D3:D{2+num_rows}')
    
    deploy_dv = create_data_validation(DEPLOYMENT_STATUS, allow_blank=False)
    ws.add_data_validation(deploy_dv)
    deploy_dv.add(f'H3:H{2+num_rows}')
    
    alert_dv = create_data_validation(ALERTING_METHOD, allow_blank=True)
    ws.add_data_validation(alert_dv)
    alert_dv.add(f'J3:J{2+num_rows}')
    
    license_dv = create_data_validation(LICENSING_MODEL, allow_blank=True)
    ws.add_data_validation(license_dv)
    license_dv.add(f'K3:K{2+num_rows}')
    
    # Conditional formatting
    ws.conditional_formatting.add(f'H3:H{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Active"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(f'H3:H{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Degraded"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'H3:H{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Offline"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    
    ws.freeze_panes = 'B3'
    
    # Add total cost formula at bottom
    total_row = 3 + num_rows
    ws[f'K{total_row}'] = "TOTAL:"
    ws[f'K{total_row}'].font = Font(bold=True)
    ws[f'L{total_row}'] = f'=SUM(L3:L{2+num_rows})'
    ws[f'L{total_row}'].font = Font(bold=True)
    ws[f'L{total_row}'].number_format = '#,##0.00'
    
    return ws

def create_drift_remediation_tracking_sheet(wb, styles):
    """Create Sheet 5: Drift_Remediation_Tracking"""
    ws = wb.create_sheet("Drift_Remediation_Tracking")
    
    widths = {
        'A': 20, 'B': 25, 'C': 15, 'D': 15, 'E': 20, 'F': 15,
        'G': 25, 'H': 40, 'I': 15, 'J': 12, 'K': 25, 'L': 15,
        'M': 18, 'N': 35, 'O': 25, 'P': 18, 'Q': 15, 'R': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:R1')
    ws['A1'] = "Drift Remediation Tracking"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    headers = {
        'A': 'Drift Incident ID', 'B': 'Asset Name', 'C': 'Drift Category',
        'D': 'Detection Date', 'E': 'Remediation Owner', 'F': 'Remediation Started Date',
        'G': 'Remediation Action Taken', 'H': 'Remediation Action Detail',
        'I': 'Remediation Completed Date', 'J': 'Time to Remediate (Days)',
        'K': 'Verification Method', 'L': 'Verification Date', 'M': 'Verification Result',
        'N': 'Recurrence Prevention', 'O': 'Root Cause Remediation',
        'P': 'Remediation Status', 'Q': 'SLA Compliance', 'R': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    num_rows = 150
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col in ['J', 'P', 'Q']:
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Formulas
    for row in range(3, 3 + num_rows):
        # Column J: Time to Remediate (Days)
        ws[f'J{row}'] = f'=IF(OR(D{row}="",I{row}=""),"",I{row}-D{row})'
        
        # Column P: Remediation Status
        ws[f'P{row}'] = f'=IF(I{row}<>"","Complete",IF(J{row}="","Not Started",IF(C{row}="Critical",IF(J{row}<1,"On-Time","Overdue"),IF(C{row}="High",IF(J{row}<2,"On-Time",IF(J{row}<3,"At Risk","Overdue")),IF(J{row}<8,"On-Time","At Risk")))))'
        
        # Column Q: SLA Compliance
        ws[f'Q{row}'] = f'=IF(P{row}="Complete",IF(C{row}="Critical",IF(J{row}<=0.17,"Met","Missed"),IF(C{row}="High",IF(J{row}<=1,"Met","Missed"),IF(C{row}="Medium",IF(J{row}<=7,"Met","Missed"),"N/A"))),"In Progress")'
    
    # Data validations
    drift_cat_dv = create_data_validation(DRIFT_CATEGORY, allow_blank=False)
    ws.add_data_validation(drift_cat_dv)
    drift_cat_dv.add(f'C3:C{2+num_rows}')
    
    action_dv = create_data_validation(REMEDIATION_ACTION, allow_blank=False)
    ws.add_data_validation(action_dv)
    action_dv.add(f'G3:G{2+num_rows}')
    
    verif_method_dv = create_data_validation(VERIFICATION_METHOD, allow_blank=True)
    ws.add_data_validation(verif_method_dv)
    verif_method_dv.add(f'K3:K{2+num_rows}')
    
    verif_result_dv = create_data_validation(VERIFICATION_RESULT, allow_blank=True)
    ws.add_data_validation(verif_result_dv)
    verif_result_dv.add(f'M3:M{2+num_rows}')
    
    root_rem_dv = create_data_validation(ROOT_CAUSE_REMEDIATION, allow_blank=True)
    ws.add_data_validation(root_rem_dv)
    root_rem_dv.add(f'O3:O{2+num_rows}')
    
    # Conditional formatting
    ws.conditional_formatting.add(f'C3:C{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Critical"'], 
                   fill=PatternFill(start_color=COLORS['critical'], end_color=COLORS['critical'], fill_type='solid'),
                   font=Font(bold=True, color='FFFFFF')))
    
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Complete"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Overdue"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    
    ws.conditional_formatting.add(f'Q3:Q{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Met"'], 
                   font=Font(bold=True, color='006100')))
    ws.conditional_formatting.add(f'Q3:Q{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Missed"'], 
                   font=Font(bold=True, color='9C0006')))
    
    ws.conditional_formatting.add(f'M3:M{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Passed"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'M3:M{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Failed"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.freeze_panes = 'B3'
    protect_formula_cells(ws, 3, 2 + num_rows, ['J', 'P', 'Q'])
    
    return ws

def create_false_positive_register_sheet(wb, styles):
    """Create Sheet 6: False_Positive_Register"""
    ws = wb.create_sheet("False_Positive_Register")
    
    widths = {
        'A': 18, 'B': 15, 'C': 25, 'D': 18, 'E': 25, 'F': 40,
        'G': 25, 'H': 40, 'I': 15, 'J': 20, 'K': 25, 'L': 35,
        'M': 15, 'N': 20, 'O': 15, 'P': 25, 'Q': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:Q1')
    ws['A1'] = "False Positive Register - Alert Quality Tracking"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    headers = {
        'A': 'False Positive ID', 'B': 'Alert Date', 'C': 'Monitoring Tool',
        'D': 'Asset ID', 'E': 'Alert Type', 'F': 'Alert Message',
        'G': 'False Positive Reason', 'H': 'False Positive Detail',
        'I': 'Investigation Date', 'J': 'Investigated By', 'K': 'Tuning Action Taken',
        'L': 'Tuning Action Detail', 'M': 'Tuning Completed Date',
        'N': 'Recurrence Status', 'O': 'Last Recurrence Date',
        'P': 'False Positive Category', 'Q': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    num_rows = 75
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            cell = ws[f'{col}{row}']
            if col == 'P':
                apply_style(cell, styles['protected_cell'])
            else:
                apply_style(cell, styles['data_cell'])
    
    # Formula for False Positive Category
    for row in range(3, 3 + num_rows):
        ws[f'P{row}'] = f'=IF(OR(G{row}="Tool Misconfiguration",G{row}="Tool Bug"),"Systemic (tool issue)",IF(G{row}="Incorrect Baseline","Baseline Issue","One-Time (environmental)"))'
    
    # Data validations
    fp_reason_dv = create_data_validation(FALSE_POSITIVE_REASON, allow_blank=False)
    ws.add_data_validation(fp_reason_dv)
    fp_reason_dv.add(f'G3:G{2+num_rows}')
    
    tuning_dv = create_data_validation(TUNING_ACTION, allow_blank=False)
    ws.add_data_validation(tuning_dv)
    tuning_dv.add(f'K3:K{2+num_rows}')
    
    recur_dv = create_data_validation(RECURRENCE_STATUS, allow_blank=False)
    ws.add_data_validation(recur_dv)
    recur_dv.add(f'N3:N{2+num_rows}')
    
    # Conditional formatting
    ws.conditional_formatting.add(f'N3:N{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Not Seen Again"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'N3:N{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Recurring (Needs Further Tuning)"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Systemic (tool issue)"'], 
                   font=Font(color='9C0006')))
    ws.conditional_formatting.add(f'P3:P{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Baseline Issue"'], 
                   font=Font(color='C65911')))
    
    ws.freeze_panes = 'B3'
    protect_formula_cells(ws, 3, 2 + num_rows, ['P'])
    
    return ws

def create_monitoring_effectiveness_metrics_sheet(wb, styles):
    """Create Sheet 7: Monitoring_Effectiveness_Metrics (Dashboard)"""
    ws = wb.create_sheet("Monitoring_Effectiveness_Metrics")
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    ws.merge_cells('A1:D1')
    ws['A1'] = "Monitoring Effectiveness Metrics Dashboard"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    # SECTION A: Overall Monitoring Coverage
    row = 3
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "OVERALL MONITORING COVERAGE"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    coverage_metrics = [
        ("Total Assets in Scope", '=COUNTA(Monitoring_Coverage_Register!A3:A102)-COUNTBLANK(Monitoring_Coverage_Register!A3:A102)', "N/A", ""),
        ("Assets Monitored", '=COUNTIF(Monitoring_Coverage_Register!G3:G102,"Monitored")', "≥85%", ""),
        ("Overall Monitoring Coverage %", '=IF(B5=0,0,B6/B5*100)', "≥85%", '=IF(B7>=85,"✓ Compliant",IF(B7>=70,"\u26A0 Partial","✗ Non-Compliant"))'),
        ("Tier 1 (Critical) Coverage %", '=IF(COUNTIF(Monitoring_Coverage_Register!F3:F102,"Tier 1*")=0,0,COUNTIFS(Monitoring_Coverage_Register!F3:F102,"Tier 1*",Monitoring_Coverage_Register!G3:G102,"Monitored")/COUNTIF(Monitoring_Coverage_Register!F3:F102,"Tier 1*")*100)', "100%", '=IF(B8>=100,"✓","✗")'),
        ("Tier 2 (High) Coverage %", '=IF(COUNTIF(Monitoring_Coverage_Register!F3:F102,"Tier 2*")=0,0,COUNTIFS(Monitoring_Coverage_Register!F3:F102,"Tier 2*",Monitoring_Coverage_Register!G3:G102,"Monitored")/COUNTIF(Monitoring_Coverage_Register!F3:F102,"Tier 2*")*100)', "≥95%", '=IF(B9>=95,"✓","✗")'),
        ("Tier 3 (Standard) Coverage %", '=IF(COUNTIF(Monitoring_Coverage_Register!F3:F102,"Tier 3*")=0,0,COUNTIFS(Monitoring_Coverage_Register!F3:F102,"Tier 3*",Monitoring_Coverage_Register!G3:G102,"Monitored")/COUNTIF(Monitoring_Coverage_Register!F3:F102,"Tier 3*")*100)', "≥85%", '=IF(B10>=85,"✓","✗")'),
        ("Non-Compliant Coverage", '=COUNTIF(Monitoring_Coverage_Register!N3:N102,"Non-Compliant")', "0", '=IF(B11=0,"✓ None","✗ "&B11&" Assets")'),
        ("Monitoring Tools Active", '=COUNTIF(Monitoring_Tool_Inventory!H3:H32,"Active")', "All", ""),
        ("Tools Offline/Degraded", '=COUNTIF(Monitoring_Tool_Inventory!H3:H32,"Offline")+COUNTIF(Monitoring_Tool_Inventory!H3:H32,"Degraded")', "0", '=IF(B13=0,"✓ None","✗ "&B13&" Issues")'),
    ]
    
    row += 1
    for metric_name, value_formula, target, status_formula in coverage_metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = value_formula
        if "%" in metric_name:
            ws[f'B{row}'].number_format = '0.0'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(name='Calibri', size=11)
        ws[f'C{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        if status_formula:
            ws[f'D{row}'] = status_formula
            apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # SECTION B: Drift Detection Metrics
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "DRIFT DETECTION METRICS"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    drift_metrics = [
        ("Total Drift Incidents (All Time)", '=COUNTA(Drift_Detection_Log!A3:A152)-COUNTBLANK(Drift_Detection_Log!A3:A152)', "N/A", ""),
        ("Critical Drift Incidents", '=COUNTIF(Drift_Detection_Log!H3:H152,"Critical")', "0", '=IF(B{row}=0,"✓ None","✗ "&B{row}&" CRITICAL")'),
        ("High Drift Incidents", '=COUNTIF(Drift_Detection_Log!H3:H152,"High")', "<5/month", ""),
        ("Unauthorized Changes Detected", '=COUNTIF(Drift_Detection_Log!K3:K152,"No (Unauthorized)")', "Detect all", ""),
        ("False Positive Rate %", '=IF(B{total}=0,0,COUNTIF(Drift_Detection_Log!O3:O152,"False Positive")/B{total}*100)', "<10%", '=IF(B{row}<10,"✓ Good",IF(B{row}<20,"\u26A0 Warning","✗ Critical"))'),
    ]
    
    row += 1
    total_row = row - 4  # Reference to total incidents
    for metric_name, value_formula, target, status_formula in drift_metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        value_formula = value_formula.replace("{row}", str(row)).replace("{total}", str(total_row))
        ws[f'B{row}'] = value_formula
        if "%" in metric_name:
            ws[f'B{row}'].number_format = '0.0'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(name='Calibri', size=11)
        ws[f'C{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        status_formula = status_formula.replace("{row}", str(row)).replace("{total}", str(total_row))
        if status_formula:
            ws[f'D{row}'] = status_formula
            apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # SECTION C: Remediation Performance
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "REMEDIATION PERFORMANCE"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    remediation_metrics = [
        ("Open Drift Incidents", '=COUNTIFS(Drift_Detection_Log!O3:O152,"<>Closed",Drift_Detection_Log!O3:O152,"<>False Positive")', "Minimize", ""),
        ("Overdue Critical Drift", '=COUNTIFS(Drift_Remediation_Tracking!C3:C152,"Critical",Drift_Remediation_Tracking!P3:P152,"Overdue")', "0", '=IF(B{row}=0,"✓ None","✗ "&B{row}&" OVERDUE")'),
        ("Mean Time to Remediate (Days)", '=IFERROR(AVERAGE(Drift_Remediation_Tracking!J3:J152),0)', "<1 day", ""),
        ("SLA Compliance - Critical %", '=IF(COUNTIF(Drift_Remediation_Tracking!C3:C152,"Critical")=0,100,COUNTIFS(Drift_Remediation_Tracking!C3:C152,"Critical",Drift_Remediation_Tracking!Q3:Q152,"Met")/COUNTIF(Drift_Remediation_Tracking!C3:C152,"Critical")*100)', "100%", '=IF(B{row}>=100,"✓","✗")'),
        ("Overall Remediation Success Rate %", '=IF(COUNTA(Drift_Remediation_Tracking!M3:M152)=0,0,COUNTIF(Drift_Remediation_Tracking!M3:M152,"Passed")/COUNTA(Drift_Remediation_Tracking!M3:M152)*100)', "≥98%", '=IF(B{row}>=98,"✓ Excellent",IF(B{row}>=90,"\u26A0 Good","✗ Needs Improvement"))'),
    ]
    
    row += 1
    for metric_name, value_formula, target, status_formula in remediation_metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        value_formula = value_formula.replace("{row}", str(row))
        ws[f'B{row}'] = value_formula
        if "%" in metric_name or "Days" in metric_name:
            ws[f'B{row}'].number_format = '0.0'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(name='Calibri', size=11)
        ws[f'C{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        status_formula = status_formula.replace("{row}", str(row))
        if status_formula:
            ws[f'D{row}'] = status_formula
            apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        row += 1
    
    # Add note
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "NOTE: All metrics auto-calculated. Review monthly. Critical drift = absolute zero tolerance."
    ws[f'A{row}'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Conditional formatting for coverage percentages
    ws.conditional_formatting.add('B7:B10',
        CellIsRule(operator='greaterThanOrEqual', formula=['85'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add('B7:B10',
        CellIsRule(operator='between', formula=['70', '84.9'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add('B7:B10',
        CellIsRule(operator='lessThan', formula=['70'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    return ws

def create_coverage_gap_analysis_sheet(wb, styles):
    """Create Sheet 8: Coverage_Gap_Analysis (Dashboard)"""
    ws = wb.create_sheet("Coverage_Gap_Analysis")
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    
    ws.merge_cells('A1:E1')
    ws['A1'] = "Monitoring Coverage Gap Analysis"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    # SECTION A: Coverage by Asset Category
    row = 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "COVERAGE BY ASSET CATEGORY"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Asset Category"
    ws[f'B{row}'] = "Total Assets"
    ws[f'C{row}'] = "Monitored"
    ws[f'D{row}'] = "Coverage %"
    ws[f'E{row}'] = "Gap Priority"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    categories = ["Infrastructure", "Endpoint", "Network Services", "Applications", "Cloud", "IoT/OT"]
    row += 1
    for category in categories:
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = f'=COUNTIF(Monitoring_Coverage_Register!$D$3:$D$102,"{category}")'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = f'=COUNTIFS(Monitoring_Coverage_Register!$D$3:$D$102,"{category}",Monitoring_Coverage_Register!$G$3:$G$102,"Monitored")'
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = f'=IF(B{row}=0,0,C{row}/B{row}*100)'
        ws[f'D{row}'].number_format = '0.0'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        ws[f'E{row}'] = f'=IF(D{row}<70,"Critical Gap",IF(D{row}<85,"High Priority","Acceptable"))'
        apply_style(ws[f'E{row}'], styles['protected_cell'])
        
        row += 1
    
    # Conditional formatting
    start_row = row - len(categories)
    ws.conditional_formatting.add(f'D{start_row}:D{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['85'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'D{start_row}:D{row-1}',
        CellIsRule(operator='between', formula=['70', '84.9'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'D{start_row}:D{row-1}',
        CellIsRule(operator='lessThan', formula=['70'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    ws.conditional_formatting.add(f'E{start_row}:E{row-1}',
        CellIsRule(operator='equal', formula=['"Critical Gap"'], 
                   font=Font(bold=True, color='9C0006')))
    ws.conditional_formatting.add(f'E{start_row}:E{row-1}',
        CellIsRule(operator='equal', formula=['"High Priority"'], 
                   font=Font(bold=True, color='C65911')))
    
    # SECTION B: Coverage by Criticality
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "COVERAGE BY ASSET CRITICALITY"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Criticality"
    ws[f'B{row}'] = "Total Assets"
    ws[f'C{row}'] = "Monitored"
    ws[f'D{row}'] = "Coverage %"
    ws[f'E{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    criticalities = [
        ("Critical", "100%"),
        ("High", "≥95%"),
        ("Medium", "≥85%"),
        ("Low", "≥70%")
    ]
    
    row += 1
    for crit, target in criticalities:
        ws[f'A{row}'] = crit
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = f'=COUNTIF(Monitoring_Coverage_Register!$E$3:$E$102,"{crit}")'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = f'=COUNTIFS(Monitoring_Coverage_Register!$E$3:$E$102,"{crit}",Monitoring_Coverage_Register!$G$3:$G$102,"Monitored")'
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = f'=IF(B{row}=0,0,C{row}/B{row}*100)'
        ws[f'D{row}'].number_format = '0.0'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        if crit == "Critical":
            ws[f'E{row}'] = f'=IF(D{row}>=100,"✓","✗")'
        elif crit == "High":
            ws[f'E{row}'] = f'=IF(D{row}>=95,"✓","✗")'
        elif crit == "Medium":
            ws[f'E{row}'] = f'=IF(D{row}>=85,"✓","✗")'
        else:
            ws[f'E{row}'] = f'=IF(D{row}>=70,"✓","✗")'
        apply_style(ws[f'E{row}'], styles['protected_cell'])
        
        row += 1
    
    return ws

def create_drift_trend_analysis_sheet(wb, styles):
    """Create Sheet 9: Drift_Trend_Analysis (Dashboard)"""
    ws = wb.create_sheet("Drift_Trend_Analysis")
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    
    ws.merge_cells('A1:F1')
    ws['A1'] = "Configuration Drift Trend Analysis"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    # Note: Simplified dashboard - in real implementation would include time-series analysis
    row = 3
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "DRIFT DISTRIBUTION BY CATEGORY"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    row += 1
    ws[f'A{row}'] = "Drift Category"
    ws[f'B{row}'] = "Total"
    ws[f'C{row}'] = "Open"
    ws[f'D{row}'] = "Closed"
    ws[f'E{row}'] = "Percentage"
    ws[f'F{row}'] = "Trend"
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    categories = ["Critical", "High", "Medium", "Low", "Informational"]
    row += 1
    for category in categories:
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        
        ws[f'B{row}'] = f'=COUNTIF(Drift_Detection_Log!$H$3:$H$152,"{category}")'
        apply_style(ws[f'B{row}'], styles['protected_cell'])
        
        ws[f'C{row}'] = f'=COUNTIFS(Drift_Detection_Log!$H$3:$H$152,"{category}",Drift_Detection_Log!$O$3:$O$152,"<>Closed",Drift_Detection_Log!$O$3:$O$152,"<>False Positive")'
        apply_style(ws[f'C{row}'], styles['protected_cell'])
        
        ws[f'D{row}'] = f'=COUNTIFS(Drift_Detection_Log!$H$3:$H$152,"{category}",Drift_Detection_Log!$O$3:$O$152,"Closed")'
        apply_style(ws[f'D{row}'], styles['protected_cell'])
        
        total_incidents = 5  # Reference row for total (to be calculated)
        ws[f'E{row}'] = f'=IF(SUM(B$5:B$9)=0,0,B{row}/SUM(B$5:B$9)*100)'
        ws[f'E{row}'].number_format = '0.0'
        apply_style(ws[f'E{row}'], styles['protected_cell'])
        
        ws[f'F{row}'] = "Monitor"
        ws[f'F{row}'].font = Font(name='Calibri', size=11)
        
        row += 1
    
    # Add note
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "NOTE: Use this data to identify drift patterns. High Critical/High drift = environmental instability."
    ws[f'A{row}'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')

    return ws

def create_evidence_register_sheet(wb, styles):
    """Create Sheet 10: Evidence_Register"""
    ws = wb.create_sheet("Evidence_Register")
    
    widths = {
        'A': 18, 'B': 20, 'C': 40, 'D': 20, 'E': 25, 'F': 15,
        'G': 40, 'H': 20, 'I': 15, 'J': 15, 'K': 15, 'L': 18, 'M': 25, 'N': 35
    }
    set_column_widths(ws, widths)
    
    ws.merge_cells('A1:N1')
    ws['A1'] = "Evidence Register - Configuration Monitoring Assessment"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    headers = {
        'A': 'Evidence ID', 'B': 'Evidence Type', 'C': 'Evidence Description',
        'D': 'Related Asset(s)', 'E': 'Related Drift Incident ID(s)', 'F': 'Evidence Date',
        'G': 'Evidence Location', 'H': 'Evidence Owner', 'I': 'Evidence Classification',
        'J': 'Retention Period', 'K': 'Last Verified Date', 'L': 'Verification Status',
        'M': 'Linked Control Requirement', 'N': 'Notes'
    }
    
    for col, header in headers.items():
        cell = ws[f'{col}2']
        cell.value = header
        apply_style(cell, styles['column_header'])
    ws.row_dimensions[2].height = 30
    
    num_rows = 100
    for row in range(3, 3 + num_rows):
        for col in headers.keys():
            apply_style(ws[f'{col}{row}'], styles['data_cell'])
    
    # Data validations
    evid_type_dv = create_data_validation(EVIDENCE_TYPE, allow_blank=False)
    ws.add_data_validation(evid_type_dv)
    evid_type_dv.add(f'B3:B{2+num_rows}')
    
    classif_dv = create_data_validation(EVIDENCE_CLASSIFICATION, allow_blank=False)
    ws.add_data_validation(classif_dv)
    classif_dv.add(f'I3:I{2+num_rows}')
    
    retention_dv = create_data_validation(RETENTION_PERIOD, allow_blank=False)
    ws.add_data_validation(retention_dv)
    retention_dv.add(f'J3:J{2+num_rows}')
    
    verif_dv = create_data_validation(EVIDENCE_VERIFICATION, allow_blank=False)
    ws.add_data_validation(verif_dv)
    verif_dv.add(f'L3:L{2+num_rows}')
    
    # Conditional formatting
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Verified"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'L3:L{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Missing"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid'),
                   font=Font(bold=True)))
    
    ws.conditional_formatting.add(f'I3:I{2+num_rows}',
        CellIsRule(operator='equal', formula=['"Restricted"'], 
                   font=Font(color='9C0006')))
    
    ws.freeze_panes = 'B3'
    
    return ws

def create_approval_signoff_sheet(wb, styles):
    """Create Sheet 11: Approval_Sign_Off"""
    ws = wb.create_sheet("Approval_Sign_Off")
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 60
    
    ws.merge_cells('A1:B1')
    ws['A1'] = "Assessment Approval Sign-Off"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 35
    
    row = 3
    
    # SECTION A: Document Information
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "DOCUMENT INFORMATION"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    doc_info = [
        ("Assessment Title", "Configuration Monitoring Assessment - Control A.8.9"),
        ("Assessment Period", "[Start Date DD.MM.YYYY] to [End Date DD.MM.YYYY]"),
        ("Document ID", DOCUMENT_ID),
        ("Version", WORKBOOK_VERSION),
        ("Assessment Date", datetime.now().strftime('%d.%m.%Y')),
    ]
    
    row += 1
    for field, value in doc_info:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = Font(name='Calibri', size=11)
        row += 1
    
    # SECTION B: Preparer Sign-Off
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "PREPARER SIGN-OFF"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    preparer_fields = [
        ("Preparer Name", ""),
        ("Preparer Role", ""),
        ("Preparer Signature", ""),
        ("Date Prepared", ""),
        ("Completeness Attestation", 
         "I attest that monitoring coverage has been documented accurately and drift incidents have been tracked completely."),
    ]
    
    row += 1
    for field, default_value in preparer_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws[f'B{row}'] = default_value
        if "Attestation" in field:
            ws[f'B{row}'].font = Font(name='Calibri', size=10, italic=True)
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['info_bg'], end_color=COLORS['info_bg'], fill_type='solid')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 40
        else:
            ws[f'B{row}'].font = Font(name='Calibri', size=11)
        row += 1
    
    # SECTION C: Reviewer Sign-Off
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "REVIEWER SIGN-OFF"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    reviewer_fields = [
        ("Reviewer Name", ""),
        ("Reviewer Role", ""),
        ("Reviewer Signature", ""),
        ("Date Reviewed", ""),
        ("Review Findings", ""),
        ("Gaps Identified", ""),
        ("Review Attestation", 
         "I have reviewed monitoring effectiveness and verified drift detection capabilities. Coverage gaps and remediation improvements have been identified."),
    ]
    
    row += 1
    for field, default_value in reviewer_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws[f'B{row}'] = default_value
        if "Attestation" in field:
            ws[f'B{row}'].font = Font(name='Calibri', size=10, italic=True)
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['info_bg'], end_color=COLORS['info_bg'], fill_type='solid')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 40
        else:
            ws[f'B{row}'].font = Font(name='Calibri', size=11)
            if field in ["Review Findings", "Gaps Identified"]:
                ws.row_dimensions[row].height = 30
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        row += 1
    
    # SECTION D: Approver Sign-Off
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "APPROVER SIGN-OFF"
    apply_style(ws[f'A{row}'], styles['section_header'])
    ws.row_dimensions[row].height = 25
    
    approver_fields = [
        ("Approver Name", ""),
        ("Approver Role", ""),
        ("Approver Signature", ""),
        ("Date Approved", ""),
        ("Approval Decision", ""),
        ("Conditions/Comments", ""),
        ("Next Assessment Due", ""),
        ("Approver Attestation", 
         "I approve this monitoring assessment and authorize budget for monitoring tool expansion and gap remediation."),
    ]
    
    row += 1
    decision_row = row + 4
    for field, default_value in approver_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws[f'B{row}'] = default_value
        if "Attestation" in field:
            ws[f'B{row}'].font = Font(name='Calibri', size=10, italic=True)
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['info_bg'], end_color=COLORS['info_bg'], fill_type='solid')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 40
        else:
            ws[f'B{row}'].font = Font(name='Calibri', size=11)
            if field == "Conditions/Comments":
                ws.row_dimensions[row].height = 30
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        if field == "Approval Decision":
            approval_dv = create_data_validation(APPROVAL_DECISION, allow_blank=True)
            ws.add_data_validation(approval_dv)
            approval_dv.add(f'B{row}')
        
        row += 1
    
    # Conditional formatting for Approval Decision
    ws.conditional_formatting.add(f'B{decision_row}:B{decision_row}',
        CellIsRule(operator='equal', formula=['"Approved"'], 
                   fill=PatternFill(start_color=COLORS['compliant'], end_color=COLORS['compliant'], fill_type='solid')))
    ws.conditional_formatting.add(f'B{decision_row}:B{decision_row}',
        CellIsRule(operator='equal', formula=['"Approved with Conditions"'], 
                   fill=PatternFill(start_color=COLORS['partial'], end_color=COLORS['partial'], fill_type='solid')))
    ws.conditional_formatting.add(f'B{decision_row}:B{decision_row}',
        CellIsRule(operator='equal', formula=['"Not Approved - Revisions Required"'], 
                   fill=PatternFill(start_color=COLORS['non_compliant'], end_color=COLORS['non_compliant'], fill_type='solid')))
    
    return ws

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main function to generate the configuration monitoring assessment workbook."""
    print("=" * 70)
    print(f"Generating {WORKBOOK_TITLE} Workbook")
    print("=" * 70)
    print(f"Document ID: {DOCUMENT_ID}")
    print(f"Version: {WORKBOOK_VERSION}")
    print(f"Date: {datetime.now().strftime('%d.%m.%Y')}")
    print("-" * 70)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create styles
    styles = create_styles()
    
    # Create all sheets
    print("Creating sheets...")
    
    print("  1/12 Creating Lookup_Tables (hidden)...")
    create_lookup_tables(wb, styles)
    
    print("  2/12 Creating Instructions sheet...")
    create_instructions_sheet(wb, styles)
    
    print("  3/12 Creating Monitoring_Coverage_Register sheet (100 rows)...")
    create_monitoring_coverage_register_sheet(wb, styles)
    
    print("  4/12 Creating Drift_Detection_Log sheet (150 rows)...")
    create_drift_detection_log_sheet(wb, styles)
    
    print("  5/12 Creating Monitoring_Tool_Inventory sheet (30 rows)...")
    create_monitoring_tool_inventory_sheet(wb, styles)
    
    print("  6/12 Creating Drift_Remediation_Tracking sheet (150 rows)...")
    create_drift_remediation_tracking_sheet(wb, styles)
    
    print("  7/12 Creating False_Positive_Register sheet (75 rows)...")
    create_false_positive_register_sheet(wb, styles)
    
    print("  8/12 Creating Monitoring_Effectiveness_Metrics sheet (dashboard)...")
    create_monitoring_effectiveness_metrics_sheet(wb, styles)
    
    print("  9/12 Creating Coverage_Gap_Analysis sheet (dashboard)...")
    create_coverage_gap_analysis_sheet(wb, styles)
    
    print(" 10/12 Creating Drift_Trend_Analysis sheet (dashboard)...")
    create_drift_trend_analysis_sheet(wb, styles)
    
    print(" 11/12 Creating Evidence_Register sheet (100 rows)...")
    create_evidence_register_sheet(wb, styles)
    
    print(" 12/12 Creating Approval_Sign_Off sheet...")
    create_approval_signoff_sheet(wb, styles)
    
    print("  ✓ All sheets created successfully")
    
    # Set workbook properties
    wb.properties.title = WORKBOOK_TITLE
    wb.properties.subject = f"ISMS Control A.8.9 - {WORKBOOK_TITLE}"
    wb.properties.creator = "[Organization] ISMS Implementation Team"
    wb.properties.created = datetime.now()
    wb.properties.description = "Assessment workbook for ISO 27001:2022 Control A.8.9 Configuration Monitoring requirements"
    
    # Save workbook
    print("-" * 70)
    print("Saving workbook...")
    wb.save(FILENAME)
    
    print("=" * 70)
    print("✓ Workbook generated successfully!")
    print("=" * 70)
    print(f"Output File: {FILENAME}")
    print(f"File Size: {os.path.getsize(FILENAME) / 1024:.1f} KB")
    print(f"Total Sheets: 12 (11 visible + 1 hidden lookup table)")
    print("-" * 70)
    print("\nWorkbook Structure:")
    print("  1.  Instructions - Usage guidance, monitoring methods, drift categories")
    print("  2.  Monitoring_Coverage_Register - 100 rows for asset monitoring inventory")
    print("  3.  Drift_Detection_Log - 150 rows for drift incident records")
    print("  4.  Monitoring_Tool_Inventory - 30 rows for tool capabilities")
    print("  5.  Drift_Remediation_Tracking - 150 rows for remediation actions")
    print("  6.  False_Positive_Register - 75 rows for alert quality tracking")
    print("  7.  Monitoring_Effectiveness_Metrics - Auto-calculated KPI dashboard")
    print("  8.  Coverage_Gap_Analysis - Coverage analysis by category/criticality")
    print("  9.  Drift_Trend_Analysis - Temporal drift pattern analysis")
    print("  10. Evidence_Register - 100 rows for evidence documentation")
    print("  11. Approval_Sign_Off - Three-tier approval signatures")
    print("  12. Lookup_Tables (hidden) - 43-type asset taxonomy")
    print("-" * 70)
    print("\nKey Monitoring Metrics:")
    print("\u2022 Coverage: Tier 1=100%, Tier 2≥95%, Tier 3≥85%, Tier 4≥70%")
    print("\u2022 MTTD (Mean Time to Detect): <1 hour (Tier 1), <24 hours (Tier 2)")
    print("\u2022 MTTR (Mean Time to Remediate): <4 hours (Critical), <1 day (High)")
    print("\u2022 False Positive Rate: <10%")
    print("\u2022 Critical Drift: ZERO TOLERANCE (immediate escalation)")
    print("\u2022 SLA Compliance: 100% (Critical), ≥95% (High)")
    print("-" * 70)
    print("\nNext Steps:")
    print("1. Open workbook in Excel/LibreOffice")
    print("2. Verify all sheets, validations, and formulas")
    print("3. Review Instructions for monitoring methods and drift categories")
    print("4. Customize dropdown values if needed (see CONFIGURATION section)")
    print("5. Document current monitoring coverage")
    print("6. Establish drift detection logging process")
    print("7. Review dashboards monthly for trends")
    print("8. Address coverage gaps identified in Coverage_Gap_Analysis")
    print("-" * 70)
    print("\nIMPORTANT REMINDERS:")
    print("\u2022 This is a SAMPLE workbook - customize for your environment")
    print("\u2022 Critical drift requires immediate action (<4 hours)")
    print("\u2022 Log all drift incidents even if immediately remediated")
    print("\u2022 False positives must be analyzed and tuning documented")
    print("\u2022 Tools marked 'Offline' create monitoring gaps")
    print("\u2022 Integration with A.8.9.1 (baselines) and A.8.9.2 (changes) is critical")
    print("\u2022 Protected cells (gray) contain formulas - do not edit")
    print("\u2022 Retain workbook and evidence for audit (minimum 3 years)")
    print("=" * 70)

if __name__ == "__main__":
    main()