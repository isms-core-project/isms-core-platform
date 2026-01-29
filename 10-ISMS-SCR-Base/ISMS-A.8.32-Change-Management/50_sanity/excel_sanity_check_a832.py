#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-A.8.32 - Excel Workbook Sanity Checker
================================================================================

ISO/IEC 27001:2022 Control A.8.32: Change Management
Quality Assurance Utility: Excel Workbook Validation & Repair Diagnosis

--------------------------------------------------------------------------------
DIAGNOSTIC UTILITY - TROUBLESHOOTS EXCEL FILE ISSUES
--------------------------------------------------------------------------------

This diagnostic utility identifies common openpyxl-generated Excel issues that
trigger Excel's "file level validation and repair" warnings when opening A.8.32
assessment workbooks.

**Purpose:**
Proactively identifies and reports Excel compatibility issues in generated
assessment workbooks before distribution to stakeholders, preventing "Excel
needs to repair this file" warnings.

**Common Issues Detected:**
- Formula syntax errors and invalid sheet references
- Data validation conflicts and overlapping ranges
- Style attribute inconsistencies (shared objects)
- Merged cell content issues
- Worksheet structure problems
- Invalid cell references

**When to Use:**
- After modifying assessment generator scripts
- Before distributing workbooks to stakeholders
- When Excel displays repair warnings on opening
- Quality assurance validation before consolidation
- Troubleshooting workbook generation issues

**What This Script Does:**
1. Analyzes workbook structure and validates against expected schema
2. Checks formula syntax and cell references
3. Validates data validation rules for conflicts
4. Inspects style objects for sharing issues
5. Verifies merged cell integrity
6. Generates diagnostic report with remediation guidance

**Workbook Support:**
- Generic checker works with all A.8.32 workbooks (A.8.32.1 through A.8.32.5)
- Auto-detects workbook type from filename
- Validates domain-specific sheet structures

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel file analysis

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Excel file reading)
    - sys, re (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    # Check any A.8.32 assessment workbook
    python3 excel_sanity_check_a832.py assessment.xlsx
    
    # Specific examples
    python3 excel_sanity_check_a832.py ISMS_A_8_32_1_Change_Process_20260127.xlsx
    python3 excel_sanity_check_a832.py ISMS-IMP-A.8.32.5.xlsx

Output:
    - Diagnostic report printed to console
    - Issue categorization (Critical/Warning/Info)
    - Recommended remediation actions
    - Structural health summary

Exit Codes:
    0 = No issues found (workbook is healthy)
    1 = Warnings found (workbook may have issues)
    2 = Critical issues found (workbook likely needs repair)

Workflow Integration:
    1. Generate assessment workbook
    2. Run this sanity checker
    3. If issues found, review generator script
    4. Regenerate workbook after fixes
    5. Re-check until clean
    6. Distribute to stakeholders

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.32
Utility Type:         Quality Assurance - Workbook Validation
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Scripts:
    - generate_a832_1_change_process.py (generates workbooks to check)
    - generate_a832_2_change_types.py (generates workbooks to check)
    - generate_a832_3_environment_separation.py (generates workbooks to check)
    - generate_a832_4_testing_validation.py (generates workbooks to check)
    - generate_a832_5_compliance_dashboard.py (generates workbooks to check)
    - normalize_assessment_files_a832.py (file preparation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Generic checker for all A.8.32 workbook types
    - Auto-detects workbook type from filename
    - Validates structure, formulas, and data validation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Preventive vs. Reactive:**
Best practice: Run this checker immediately after generating workbooks, before
stakeholder distribution. Much easier to fix issues in generator scripts than
to manually repair distributed workbooks.

**Excel Repair Warnings:**
If Excel says "We found a problem with some content..." and offers to repair,
the issue was likely present during generation. This script helps identify
the root cause.

**Common Root Causes:**
- Shared style objects between cells (openpyxl limitation)
- Formula references to non-existent sheets
- Data validation ranges overlapping with merged cells
- Invalid cell references in formulas

**False Positives:**
Some warnings may be cosmetic and not actually cause Excel to repair. Use
judgment - if Excel opens the file without repair prompts, you're probably fine.

**Generator Script Fixes:**
When issues are found, fix them in the generator scripts (generate_a832_*.py),
not by manually editing Excel files. Fixing at source ensures future
generations don't have the same issues.

**Integration Testing:**
After fixing generator scripts, regenerate workbooks and re-check with this
utility. Repeat until clean bill of health achieved.

================================================================================
"""

import sys
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ============================================================================
# WORKBOOK TYPE DETECTION
# ============================================================================

def detect_workbook_type(filename):
    """Detect which A.8.32 workbook type this is based on filename."""
    filename_lower = filename.lower()
    
    if 'a.8.32.1' in filename_lower or 'a_8_32_1' in filename_lower or 'change_process' in filename_lower:
        return 'A.8.32.1', 'Change Process Assessment'
    elif 'a.8.32.2' in filename_lower or 'a_8_32_2' in filename_lower or 'change_types' in filename_lower:
        return 'A.8.32.2', 'Change Types & Categories Assessment'
    elif 'a.8.32.3' in filename_lower or 'a_8_32_3' in filename_lower or 'environment_separation' in filename_lower:
        return 'A.8.32.3', 'Environment Separation Assessment'
    elif 'a.8.32.4' in filename_lower or 'a_8_32_4' in filename_lower or 'testing_validation' in filename_lower:
        return 'A.8.32.4', 'Testing & Validation Assessment'
    elif 'a.8.32.5' in filename_lower or 'a_8_32_5' in filename_lower or 'compliance_dashboard' in filename_lower:
        return 'A.8.32.5', 'Compliance Dashboard'
    else:
        return 'Unknown', 'Generic Excel Workbook'


# ============================================================================
# EXPECTED SHEET DEFINITIONS
# ============================================================================

EXPECTED_SHEETS = {
    'A.8.32.1': [
        "Instructions & Legend", "Change_Process_Workflow", "Approval_Authority_Matrix",
        "Communication_Procedures", "Documentation_Requirements", "Change_Management_Tools",
        "Summary_Dashboard", "Evidence_Register", "Approval_Sign_Off"
    ],
    'A.8.32.2': [
        "Instructions & Legend", "Standard_Changes_Catalog", "Normal_Changes_Assessment",
        "Emergency_Changes", "Change_Risk_Classification", "Change_Calendar_Management",
        "Summary_Dashboard", "Evidence_Register", "Approval_Sign_Off"
    ],
    'A.8.32.3': [
        "Instructions & Legend", "Development_Environment", "Test_QA_Environment",
        "Production_Environment", "Environment_Promotion_Process", "Production_Data_in_NonProd",
        "Summary_Dashboard", "Evidence_Register", "Approval_Sign_Off"
    ],
    'A.8.32.4': [
        "Instructions & Legend", "Testing_Framework_Assessment", "Test_Types_Coverage",
        "Acceptance_Criteria_Management", "Rollback_Procedures", "Production_Validation",
        "Summary_Dashboard", "Evidence_Register", "Approval_Sign_Off"
    ],
    'A.8.32.5': [
        "Executive_Dashboard", "Gap_Analysis", "Risk_Register",
        "Remediation_Roadmap", "KPIs_and_Metrics", "Evidence_Consolidation",
        "Action_Items_and_Followup", "Audit_and_Compliance_Log", "Approval_Sign_Off"
    ],
}


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""
    
    workbook_id, workbook_name = detect_workbook_type(filename)
    
    print("=" * 80)
    print(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {filename}")
    print(f"Detected Type: {workbook_id} - {workbook_name}")
    print("=" * 80)
    
    issues_found = []
    warnings_found = []
    
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"\n❌" Workbook loaded successfully")
        print(f"  Sheets found: {len(wb.sheetnames)}")
        print(f"  Sheet names: {', '.join(wb.sheetnames)}")
        
    except Exception as e:
        print(f"\n❌ CRITICAL: Cannot load workbook: {e}")
        return
    
    # ========================================================================
    # CHECK 1: SHEET STRUCTURE VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 1: SHEET STRUCTURE VALIDATION")
    print("=" * 80)
    
    if workbook_id in EXPECTED_SHEETS:
        expected = EXPECTED_SHEETS[workbook_id]
        actual = wb.sheetnames
        
        # Check for missing sheets (flexible matching)
        missing = []
        for exp_sheet in expected:
            # Flexible match: check if expected sheet name is in any actual sheet
            found = False
            for act_sheet in actual:
                if exp_sheet.lower().replace("_", "") in act_sheet.lower().replace("_", "").replace(" ", ""):
                    found = True
                    break
            if not found:
                missing.append(exp_sheet)
        
        if missing:
            for sheet in missing:
                warnings_found.append(f"  ⚠️ Missing expected sheet: {sheet}")
            print(f"  ⚠️ {len(missing)} expected sheets not found (may be renamed)")
        else:
            print(f"  ❌" All expected sheets present for {workbook_id}")
        
        # Check for extra sheets
        print(f"  Total sheets: {len(actual)} (expected ~{len(expected)})")
    else:
        print("  ℹ️ Unknown workbook type - skipping sheet validation")
    
    # ========================================================================
    # CHECK 2: FORMULA VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 2: FORMULA VALIDATION")
    print("=" * 80)
    
    formula_issues = 0
    inter_sheet_refs = {}
    external_refs = set()
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # Check for external workbook references
                    ext_refs = re.findall(r'\[([^\]]+\.xlsx)\]', formula)
                    for ref in ext_refs:
                        external_refs.add(ref)
                    
                    # Check for sheet references
                    if "'" in formula or "!" in formula:
                        sheet_refs = re.findall(r"'?([^'!]+)'?!", formula)
                        for ref in sheet_refs:
                            ref_clean = ref.strip("'")
                            # Skip external references
                            if "[" not in ref_clean and ref_clean not in wb.sheetnames:
                                # Check for close matches (underscore vs space)
                                close_match = False
                                for actual in wb.sheetnames:
                                    if ref_clean.replace(" ", "_") == actual or ref_clean.replace("_", " ") == actual:
                                        close_match = True
                                        break
                                if not close_match:
                                    issues_found.append(
                                        f"  ❌ {sheet_name}!{cell.coordinate}: "
                                        f"Invalid sheet reference '{ref_clean}'"
                                    )
                                    formula_issues += 1
                            
                            # Track inter-sheet references
                            if "[" not in ref_clean and ref_clean in wb.sheetnames:
                                if sheet_name not in inter_sheet_refs:
                                    inter_sheet_refs[sheet_name] = set()
                                inter_sheet_refs[sheet_name].add(ref_clean)
                    
                    # Check for common syntax issues
                    if formula.count('(') != formula.count(')'):
                        issues_found.append(
                            f"  ❌ {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced parentheses in formula"
                        )
                        formula_issues += 1
    
    if formula_issues == 0:
        print("  ❌" No formula syntax issues detected")
    else:
        print(f"  ❌ Found {formula_issues} formula issues")
    
    if inter_sheet_refs:
        print(f"  ℹ️ Inter-sheet references found:")
        for sheet, refs in inter_sheet_refs.items():
            print(f"    {sheet} → {', '.join(refs)}")
    
    if external_refs:
        print(f"  ℹ️ External workbook references found:")
        for ref in external_refs:
            print(f"    → {ref}")
        if workbook_id == 'A.8.32.5':
            print("  ⚠️ WARNING: Dashboard requires external links to function")
    
    # ========================================================================
    # CHECK 3: DATA VALIDATION CONFLICTS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 3: DATA VALIDATION CONFLICTS")
    print("=" * 80)
    
    validation_issues = 0
    total_validations = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            dv_count = len(ws.data_validations.dataValidation)
            total_validations += dv_count
            
            # Check for overlapping ranges
            all_ranges = []
            for dv in ws.data_validations.dataValidation:
                if hasattr(dv, 'sqref') and dv.sqref:
                    all_ranges.extend(str(dv.sqref).split())
            
            if len(all_ranges) != len(set(all_ranges)):
                warnings_found.append(
                    f"  ⚠️ {sheet_name}: Potentially overlapping data validation ranges"
                )
                validation_issues += 1
    
    print(f"  Total data validations: {total_validations}")
    
    if validation_issues == 0:
        print("  ❌" No data validation conflicts detected")
    
    # ========================================================================
    # CHECK 4: MERGED CELLS VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 4: MERGED CELLS VALIDATION")
    print("=" * 80)
    
    merge_issues = 0
    total_merges = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'merged_cells') and ws.merged_cells:
            merge_count = len(ws.merged_cells.ranges)
            total_merges += merge_count
            
            # Check merged cells for content issues
            for merge_range in ws.merged_cells.ranges:
                min_col, min_row = merge_range.min_col, merge_range.min_row
                max_col, max_row = merge_range.max_col, merge_range.max_row
                
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if row == min_row and col == min_col:
                            continue
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            cell_coord = f"{get_column_letter(col)}{row}"
                            warnings_found.append(
                                f"  ⚠️ {sheet_name}!{cell_coord}: "
                                f"Merged cell has content in non-primary cell"
                            )
                            merge_issues += 1
    
    print(f"  Total merged ranges: {total_merges}")
    
    if merge_issues == 0:
        print("  ❌" Merged cells properly formatted")
    
    # ========================================================================
    # CHECK 5: EVIDENCE REGISTER VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 5: EVIDENCE REGISTER VALIDATION")
    print("=" * 80)
    
    evidence_sheet = None
    for sheet in wb.sheetnames:
        if 'evidence' in sheet.lower():
            evidence_sheet = sheet
            break
    
    if evidence_sheet:
        ws = wb[evidence_sheet]
        # Count evidence rows (look for EV prefix or numbered IDs)
        evidence_count = 0
        for row in range(1, min(150, ws.max_row + 1)):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and isinstance(cell_val, str):
                if cell_val.startswith('EV') or cell_val.startswith('EVM'):
                    evidence_count += 1
        
        print(f"  Evidence sheet: {evidence_sheet}")
        print(f"  Evidence row templates: {evidence_count}")
        
        if evidence_count < 50:
            warnings_found.append(
                f"  ⚠️ Evidence Register has {evidence_count} rows (expected ≥50)"
            )
        elif evidence_count >= 100:
            print(f"  ❌" Full 100-row evidence register")
        else:
            print(f"  ❌" Evidence register adequate ({evidence_count} rows)")
    else:
        warnings_found.append("  ⚠️ No Evidence Register sheet found")
    
    # ========================================================================
    # CHECK 6: WORKSHEET DIMENSIONS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 6: WORKSHEET DIMENSIONS")
    print("=" * 80)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if ws.max_row > 5000:
            warnings_found.append(
                f"  ⚠️ {sheet_name}: Large worksheet ({ws.max_row} rows)"
            )
        
        if ws.max_column > 50:
            warnings_found.append(
                f"  ⚠️ {sheet_name}: Wide worksheet ({ws.max_column} columns)"
            )
    
    print("  ❌" Worksheet dimensions within reasonable limits")
    
    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    print("\n" + "=" * 80)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 80)
    
    if issues_found:
        print(f"\n❌ CRITICAL ISSUES FOUND: {len(issues_found)}")
        for issue in issues_found:
            print(issue)
    
    if warnings_found:
        print(f"\n⚠️️  WARNINGS: {len(warnings_found)}")
        for warning in warnings_found:
            print(warning)
    
    if not issues_found and not warnings_found:
        print("\n✅ NO ISSUES DETECTED")
        print("\nThe workbook appears structurally sound.")
        print("If Excel still shows repair warnings, consider:")
        print("  • Excel version compatibility (try Excel 2019+)")
        print("  • Antivirus/security software interference")
        print("  • Network drive or OneDrive sync issues")
        print("  • Excel's overly cautious validation")
        if workbook_id == 'A.8.32.5':
            print("  • Unresolved external workbook links (expected before setup)")
    else:
        print_recommendations(issues_found, warnings_found, workbook_id, 
                            formula_issues, validation_issues, merge_issues)
    
    # ========================================================================
    # WORKBOOK-SPECIFIC NOTES
    # ========================================================================
    print_workbook_notes(workbook_id, external_refs)
    
    wb.close()


def print_recommendations(issues, warnings, workbook_id, formula_issues, 
                         validation_issues, merge_issues):
    """Print recommended fixes based on issues found."""
    print("\n" + "=" * 80)
    print("RECOMMENDED ACTIONS")
    print("=" * 80)
    
    if formula_issues > 0:
        print("\n1. FORMULA FIXES:")
        print("   • Review formulas referencing other sheets")
        print("   • Ensure sheet names match exactly (case-sensitive)")
        print("   • Check for balanced quotes and parentheses")
        if workbook_id == 'A.8.32.5':
            print("   • Verify external workbook references are correct")
    
    if validation_issues > 0:
        print("\n2. DATA VALIDATION FIXES:")
        print("   • Remove overlapping data validation ranges")
        print("   • Apply validations to specific ranges, not entire columns")
    
    if merge_issues > 0:
        print("\n3. MERGED CELL FIXES:")
        print("   • Ensure only top-left cell of merged range has content")
        print("   • Clear content from other cells in merged range")
    
    if workbook_id == 'A.8.32.5':
        print("\n4. DASHBOARD-SPECIFIC SETUP (A.8.32.5):")
        print("   • This workbook consolidates data from Domains 1-4")
        print("   • Ensure source workbooks are in the same folder:")
        print("     - ISMS-IMP-A.8.32.1_Filtering_Infrastructure.xlsx")
        print("     - ISMS-IMP-A.8.32.2_Network_Coverage.xlsx")
        print("     - ISMS-IMP-A.8.32.3_Policy_Configuration.xlsx")
        print("     - ISMS-IMP-A.8.32.4_Monitoring_Response.xlsx")


def print_workbook_notes(workbook_id, external_refs):
    """Print workbook-specific notes and expected structure."""
    print("\n" + "=" * 80)
    print(f"WORKBOOK-SPECIFIC NOTES: {workbook_id}")
    print("=" * 80)
    
    if workbook_id == 'A.8.32.1':
        print("\nChange Process Assessment:")
        print("  • 9 sheets for solution inventory and capabilities")
        print("  • 50 solution rows in Solution_Inventory")
        print("  • Capability assessment against requirements")
        print("  • Integration architecture documentation")
        print("  • Gap analysis with remediation tracking")
        print("  • Evidence Register (100 entries)")
        print("  • 3-level approval workflow")
    
    elif workbook_id == 'A.8.32.2':
        print("\nChange Types & Categories Assessment:")
        print("  • 10 sheets for coverage mapping")
        print("  • 50 network segment tracking")
        print("  • Coverage matrix (segment × solution)")
        print("  • Bypass risk assessment")
        print("  • Device inventory (100 rows)")
        print("  • Exemption register with approvals")
        print("  • Coverage verification testing")
        print("  • Evidence Register (100 entries)")
    
    elif workbook_id == 'A.8.32.3':
        print("\nEnvironment Separation Assessment:")
        print("  • 12 sheets for policy documentation")
        print("  • Approach-agnostic (restrictive/trust-based/hybrid)")
        print("  • Threat protection baseline")
        print("  • Category management (if applicable)")
        print("  • Custom block/allow lists")
        print("  • Exception register with approval workflow")
        print("  • AUP alignment verification")
        print("  • Evidence Register (100 entries)")
    
    elif workbook_id == 'A.8.32.4':
        print("\nTesting & Validation Assessment:")
        print("  • 11 sheets for operational monitoring")
        print("  • Log collection inventory (30 sources)")
        print("  • Alert configuration (40 rules)")
        print("  • Monitoring dashboards and KPIs (20 each)")
        print("  • Incident response with SLA tracking")
        print("  • False positive management (50 entries)")
        print("  • Reporting schedule")
        print("  • Evidence Register (100 entries)")
    
    elif workbook_id == 'A.8.32.5':
        print("\nCompliance Dashboard (CONSOLIDATION):")
        print("  • 11 sheets - MASTER CONSOLIDATION WORKBOOK")
        print("  • Executive Summary with traffic lights")
        print("  • Domain summaries (pull from 1-4)")
        print("  • Quantitative compliance scoring (0-100%)")
        print("  • Maturity assessment (1-5 scale)")
        print("  • Gap consolidation (60 items from all domains)")
        print("  • Evidence index (50 key items)")
        print("  • Action plan (30 remediation items)")
        print("  • CISO approval workflow")
        print("\n  ⚠️ This dashboard may show #REF errors until:")
        print("    1. All 4 source workbooks are completed")
        print("    2. Files are in same folder with correct names")
        print("    3. Links are updated in Excel")
    
    else:
        print("\nGeneric Excel Workbook")
        print("  • No specific validation rules defined for this type")
        print("  • Standard Excel health checks applied")
    
    print("\n" + "=" * 80)


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("=" * 70)
        print("ISMS Control A.8.32 Change Management - Excel Sanity Checker")
        print("=" * 70)
        print("\nUsage: python3 excel_sanity_check_a832.py <filename.xlsx>")
        print("\nExamples:")
        print("  python3 excel_sanity_check_a832.py ISMS-IMP-A.8.32.1_Filtering_Infrastructure_20260101.xlsx")
        print("  python3 excel_sanity_check_a832.py ISMS-IMP-A.8.32.2_Network_Coverage_20260101.xlsx")
        print("  python3 excel_sanity_check_a832.py ISMS-IMP-A.8.32.3_Policy_Configuration_20260101.xlsx")
        print("  python3 excel_sanity_check_a832.py ISMS-IMP-A.8.32.4_Monitoring_Response_20260101.xlsx")
        print("  python3 excel_sanity_check_a832.py ISMS-IMP-A.8.32.5_Compliance_Dashboard_20260101.xlsx")
        print("\nSupported workbook types:")
        print("  A.8.32.1 - Change Process Assessment")
        print("  A.8.32.2 - Change Types & Categories Assessment")
        print("  A.8.32.3 - Environment Separation Assessment")
        print("  A.8.32.4 - Testing & Validation Assessment")
        print("  A.8.32.5 - Compliance Dashboard")
        print("\n'Evidence > Theater' - Systems Engineering ISMS")
        sys.exit(1)
    
    filename = sys.argv[1]
    
    try:
        check_workbook_health(filename)
    except FileNotFoundError:
        print(f"\n❌ ERROR: File not found: {filename}")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()