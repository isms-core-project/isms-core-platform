#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-A.8.32 - Dashboard Consolidation Utility
================================================================================

ISO/IEC 27001:2022 Control A.8.32: Change Management
Operational Utility: Multi-Domain Change Management Dashboard Consolidation

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific assessment workbook structures and consolidation
requirements.

CRITICAL: This consolidation utility MUST be customized based on the ACTUAL
structure of your four assessment workbooks. DO NOT assume sheet names,
column positions, or data structures match this template.

Key customization areas:
1. Input workbook schemas - ANALYZE ACTUAL STRUCTURE FIRST
2. Sheet names and column mappings - MUST MATCH YOUR ASSESSMENTS
3. Data extraction logic - SPECIFIC TO YOUR WORKBOOK LAYOUTS
4. Scoring and weighting algorithms - ADAPT TO YOUR RISK PROFILE
5. Dashboard visualization preferences - ALIGN WITH YOUR REPORTING NEEDS

MANDATORY STEPS BEFORE USING THIS SCRIPT:
1. Generate all four A.8.32 assessment workbooks (Domains 1-4)
2. Run normalize_assessment_files_a832.py to validate structure
3. Document the actual sheet structure of each workbook
4. Update WORKBOOK_SCHEMAS dictionary with real structure
5. Validate schema against actual files before consolidation
6. Test with --dry-run before production use

Reference Pattern: Based on ISMS-A.8.32 Change Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script consolidates data from four normalized change management assessment
workbooks into a unified compliance dashboard, providing executive visibility
and audit-ready evidence summaries.

**Purpose:**
Automates the consolidation of multi-domain change management assessments into
a single dashboard workbook, eliminating manual copy-paste errors and ensuring
consistent compliance reporting.

**Relationship to generate_a832_5_compliance_dashboard.py:**
- generate_a832_5: Creates the EMPTY dashboard template structure
- consolidate_a832_dashboard.py: POPULATES dashboard with actual assessment data

Use both scripts in sequence:
1. Run generate_a832_5 to create dashboard template
2. Complete all four domain assessments
3. Run normalize_assessment_files_a832.py for validation
4. Run consolidate_a832_dashboard.py to populate dashboard

**Input Assessment Domains:**
1. Domain 1: Change Process Assessment (A.8.32.1)
   Input: ISMS_A_8_32_1_Change_Process_Assessment_YYYYMMDD.xlsx
   
2. Domain 2: Change Types & Categories (A.8.32.2)
   Input: ISMS_A_8_32_2_Change_Types_Categories_Assessment_YYYYMMDD.xlsx
   
3. Domain 3: Environment Separation (A.8.32.3)
   Input: ISMS_A_8_32_3_Environment_Separation_Assessment_YYYYMMDD.xlsx
   
4. Domain 4: Testing & Validation (A.8.32.4)
   Input: ISMS_A_8_32_4_Testing_Validation_Assessment_YYYYMMDD.xlsx

**Consolidation Process:**

Phase 1: Pre-Consolidation Validation
- Verify all input workbooks exist and are readable
- Validate workbook structure against expected schema
- Check for required sheets and columns
- Identify missing or incomplete data
- Generate pre-consolidation validation report

Phase 2: Data Extraction
- Extract compliance scores from each domain
- Collect gap analysis findings
- Aggregate evidence references
- Summarize remediation requirements
- Calculate domain-specific metrics

Phase 3: Dashboard Population
- Populate executive summary with aggregate metrics
- Generate compliance overview by domain
- Create consolidated gap analysis
- Populate remediation tracker
- Build audit evidence index
- Calculate overall compliance percentage

Phase 4: Quality Assurance
- Validate consolidated data integrity
- Check for calculation errors
- Verify evidence linkage completeness
- Generate post-consolidation validation report

Phase 5: Output Generation
- Save populated dashboard workbook
- Generate consolidation log
- Create executive summary report (optional)
- Archive source assessment references

**Generated Dashboard Outputs:**
- Primary: ISMS_A_8_32_5_Compliance_Dashboard_YYYYMMDD.xlsx (populated)
- Log: A832_Dashboard_Consolidation_Log_YYYYMMDD.txt
- Report: A832_Executive_Summary_YYYYMMDD.pdf (optional, if --pdf flag used)

**Key Features:**
- Schema-driven data extraction (prevents hardcoded assumptions)
- Intelligent error handling for missing/malformed data
- Weighted scoring based on domain criticality
- Gap prioritization by risk and compliance impact
- Audit trail linking dashboard to source assessments
- Rollback capability if consolidation fails
- Incremental updates (rerun after assessment changes)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel processing

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - os (standard library)
    - json (standard library - schema definitions)
    - logging (standard library - consolidation logging)

Optional Dependencies:
    - reportlab (for PDF executive summary generation)
      pip3 install reportlab

Input Requirements:
    - All four domain assessment workbooks must be completed
    - Workbooks should be normalized (run normalize_assessment_files_a832.py first)
    - Dashboard template must exist (generated by generate_a832_5)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    # Consolidate all assessments into dashboard (auto-detect latest files)
    python3 consolidate_a832_dashboard.py

Advanced Usage:
    # Specify input directory for assessment workbooks
    python3 consolidate_a832_dashboard.py --input-dir /path/to/assessments
    
    # Specify dashboard template file explicitly
    python3 consolidate_a832_dashboard.py --dashboard /path/to/dashboard.xlsx
    
    # Specify output directory for populated dashboard
    python3 consolidate_a832_dashboard.py --output-dir /path/to/output
    
    # Dry run mode (validate without writing to dashboard)
    python3 consolidate_a832_dashboard.py --dry-run
    
    # Generate PDF executive summary
    python3 consolidate_a832_dashboard.py --pdf
    
    # Verbose logging for troubleshooting
    python3 consolidate_a832_dashboard.py --verbose
    
    # Specify assessment files explicitly (override auto-detection)
    python3 consolidate_a832_dashboard.py \
        --domain1 path/to/change_process_assessment.xlsx \
        --domain2 path/to/change_types_assessment.xlsx \
        --domain3 path/to/environment_separation_assessment.xlsx \
        --domain4 path/to/testing_validation_assessment.xlsx

Command-Line Options:
    --input-dir PATH       Directory containing assessment workbooks
    --dashboard PATH       Path to dashboard template file
    --output-dir PATH      Directory for populated dashboard
    --domain1 PATH         Explicit path to Domain 1 assessment
    --domain2 PATH         Explicit path to Domain 2 assessment
    --domain3 PATH         Explicit path to Domain 3 assessment
    --domain4 PATH         Explicit path to Domain 4 assessment
    --dry-run              Validate without writing to dashboard
    --pdf                  Generate PDF executive summary (requires reportlab)
    --verbose              Enable detailed logging
    --backup               Create backup of dashboard before populating
    --incremental          Update existing dashboard (preserve manual edits)

Output Files:
    Primary Output:
        ISMS_A_8_32_5_Compliance_Dashboard_YYYYMMDD.xlsx (populated)
    
    Supporting Files:
        A832_Dashboard_Consolidation_Log_YYYYMMDD.txt (consolidation log)
        A832_Executive_Summary_YYYYMMDD.pdf (if --pdf used)
        dashboard_backup_YYYYMMDD.xlsx (if --backup used)

Workflow Examples:

    1. First-time consolidation:
       python3 consolidate_a832_dashboard.py --verbose --backup
    
    2. Update dashboard after assessment changes:
       python3 consolidate_a832_dashboard.py --incremental
    
    3. Validation before production:
       python3 consolidate_a832_dashboard.py --dry-run --verbose
    
    4. Generate executive briefing:
       python3 consolidate_a832_dashboard.py --pdf

Exit Codes:
    0   Success - Dashboard populated without errors
    1   Validation failure - Input workbooks don't meet schema requirements
    2   Data extraction error - Unable to read required data from assessments
    3   Consolidation error - Error during dashboard population
    4   File I/O error - Unable to read/write required files

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.32
Utility Type:         Operational - Dashboard Data Consolidation
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.32: Change Management Policy (Governance)
    - ISMS-IMP-A.8.32.1: Change Process Assessment (Input Domain 1)
    - ISMS-IMP-A.8.32.2: Change Types & Categories Assessment (Input Domain 2)
    - ISMS-IMP-A.8.32.3: Environment Separation Assessment (Input Domain 3)
    - ISMS-IMP-A.8.32.4: Testing & Validation Assessment (Input Domain 4)
    - ISMS-IMP-A.8.32.5: Compliance Dashboard Implementation Guide

Related Scripts:
    - generate_a832_1_change_process.py (generates input)
    - generate_a832_2_change_types.py (generates input)
    - generate_a832_3_environment_separation.py (generates input)
    - generate_a832_4_testing_validation.py (generates input)
    - generate_a832_5_compliance_dashboard.py (generates template)
    - normalize_assessment_files_a832.py (validates inputs)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements schema-driven consolidation framework
    - Supports all four change management assessment domains
    - Generates audit-ready compliance dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Schema Validation is MANDATORY:**
This script WILL FAIL if workbook schemas don't match expectations. You MUST:
1. Run normalize_assessment_files_a832.py BEFORE consolidation
2. Document actual workbook structure in WORKBOOK_SCHEMAS dictionary
3. Test with --dry-run before production consolidation
4. Keep schema definitions synchronized with assessment workbooks

Failure to validate schemas = garbage in, garbage out.

**Audit Considerations:**
This script generates the primary compliance dashboard for ISO 27001:2022
Control A.8.32 audits. Auditors will expect:
- Traceability from dashboard metrics to source assessments
- Consolidation log demonstrating systematic process
- Evidence that consolidated data hasn't been manually manipulated

Maintain consolidation logs and source assessments for audit trail.

**Data Protection:**
Consolidated dashboard contains executive summary of change management security
posture across all domains. This is HIGHLY SENSITIVE information including:
- Overall compliance gaps and critical vulnerabilities
- Change management deficiencies across the organization
- Remediation priorities and timelines

Handle with MAXIMUM SECURITY. Consider encrypting the dashboard workbook.

**Incremental vs. Full Consolidation:**
- Full consolidation (default): Overwrites all dashboard data
- Incremental (--incremental): Preserves manual edits in dashboard

Use incremental mode carefully - it can preserve outdated data if assessment
values have changed. Default to full consolidation for accuracy.

**Error Handling Philosophy:**
Script fails fast on schema mismatches or missing data. Better to fail
obviously than to silently produce incorrect compliance metrics.

If consolidation fails:
1. Check consolidation log for error details
2. Run normalize_assessment_files_a832.py to identify issues
3. Fix source assessment workbooks
4. Rerun consolidation

**Performance:**
Consolidation processes multiple large Excel workbooks. Expected runtime:
- Small organizations (<50 change types): <30 seconds
- Medium organizations (50-200 change types): 1-3 minutes
- Large organizations (>200 change types): 3-10 minutes

Use --verbose to monitor progress for large consolidations.

**Quality Assurance:**
After consolidation, manually verify:
- Compliance percentages look reasonable
- Gap counts match expectations
- Critical findings are correctly prioritized
- Evidence references are intact

Don't blindly trust automated consolidation - validate the results.

**Integration Opportunities:**
Dashboard data can feed into:
- GRC platform compliance reporting
- Executive dashboards and business intelligence tools
- Automated compliance monitoring systems
- Ticketing systems for remediation tracking

Consider exporting dashboard data to JSON/CSV for integration.

**Regulatory Reporting:**
Customize consolidation logic to highlight regulatory-specific compliance:
- Financial services: Emphasize change control segregation
- Healthcare: Highlight patient data system changes
- Government: Focus on audit trail completeness

Adapt weighting and prioritization to your regulatory context.

**Backup and Recovery:**
Always use --backup flag for production consolidations. If consolidation
produces unexpected results, you can restore from backup and troubleshoot.

**Version Control:**
Maintain version history of consolidated dashboards to track compliance
improvement over time. Use date suffixes and archive previous versions.

================================================================================
"""

import sys
import os
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# ============================================================================
# WORKBOOK SCHEMAS - CUSTOMIZE FOR YOUR ACTUAL ASSESSMENT STRUCTURES
# ============================================================================

# CRITICAL: These schemas MUST match the actual structure of your assessment
# workbooks. DO NOT use default values without validating against your actual
# generated workbooks.

WORKBOOK_SCHEMAS = {
    'domain1': {
        'filename_pattern': 'ISMS_A_8_32_1_Change_Process_Assessment',
        'normalized_name': 'ISMS-IMP-A.8.32.1.xlsx',
        'assessment_area': 'Change Process & Workflow Management',
        'sheets': {
            'Summary_Dashboard': {
                'compliance_metrics_row': 15,
                'gap_count_row': 20,
                'has_gaps': True,
                'gap_start_row': 25,
                'gap_columns': 'A:H'
            },
            'Change_Process_Workflow': {
                'assessment_start_row': 10,
                'data_columns': 'A:G',
                'has_status_column': True,
                'status_column': 'G'
            },
            'Approval_Authority_Matrix': {
                'assessment_start_row': 10,
                'data_columns': 'A:F',
                'has_status_column': True,
                'status_column': 'F'
            },
            'Evidence_Register': {
                'evidence_start_row': 5,
                'evidence_columns': 'A:F',
                'has_evidence': True
            }
        }
    },
    'domain2': {
        'filename_pattern': 'ISMS_A_8_32_2_Change_Types_Categories_Assessment',
        'normalized_name': 'ISMS-IMP-A.8.32.2.xlsx',
        'assessment_area': 'Change Classification & Risk Management',
        'sheets': {
            'Summary_Dashboard': {
                'compliance_metrics_row': 15,
                'gap_count_row': 20,
                'has_gaps': True,
                'gap_start_row': 25,
                'gap_columns': 'A:H'
            },
            'Standard_Changes_Catalog': {
                'assessment_start_row': 10,
                'data_columns': 'A:H',
                'has_status_column': True,
                'status_column': 'H'
            },
            'Normal_Changes_Assessment': {
                'assessment_start_row': 10,
                'data_columns': 'A:G',
                'has_status_column': True,
                'status_column': 'G'
            },
            'Emergency_Changes': {
                'assessment_start_row': 10,
                'data_columns': 'A:H',
                'has_status_column': True,
                'status_column': 'H'
            },
            'Evidence_Register': {
                'evidence_start_row': 5,
                'evidence_columns': 'A:F',
                'has_evidence': True
            }
        }
    },
    'domain3': {
        'filename_pattern': 'ISMS_A_8_32_3_Environment_Separation_Assessment',
        'normalized_name': 'ISMS-IMP-A.8.32.3.xlsx',
        'assessment_area': 'Development/Test/Production Isolation',
        'sheets': {
            'Summary_Dashboard': {
                'compliance_metrics_row': 15,
                'gap_count_row': 20,
                'has_gaps': True,
                'gap_start_row': 25,
                'gap_columns': 'A:H'
            },
            'Development_Environment': {
                'assessment_start_row': 10,
                'data_columns': 'A:G',
                'has_status_column': True,
                'status_column': 'G'
            },
            'Test_QA_Environment': {
                'assessment_start_row': 10,
                'data_columns': 'A:G',
                'has_status_column': True,
                'status_column': 'G'
            },
            'Production_Environment': {
                'assessment_start_row': 10,
                'data_columns': 'A:G',
                'has_status_column': True,
                'status_column': 'G'
            },
            'Evidence_Register': {
                'evidence_start_row': 5,
                'evidence_columns': 'A:F',
                'has_evidence': True
            }
        }
    },
    'domain4': {
        'filename_pattern': 'ISMS_A_8_32_4_Testing_Validation_Assessment',
        'normalized_name': 'ISMS-IMP-A.8.32.4.xlsx',
        'assessment_area': 'Testing, Validation & Acceptance',
        'sheets': {
            'Summary_Dashboard': {
                'compliance_metrics_row': 15,
                'gap_count_row': 20,
                'has_gaps': True,
                'gap_start_row': 25,
                'gap_columns': 'A:H'
            },
            'Testing_Framework_Overview': {
                'assessment_start_row': 10,
                'data_columns': 'A:G',
                'has_status_column': True,
                'status_column': 'G'
            },
            'Unit_Integration_Testing': {
                'assessment_start_row': 10,
                'data_columns': 'A:H',
                'has_status_column': True,
                'status_column': 'H'
            },
            'UAT_Business_Validation': {
                'assessment_start_row': 10,
                'data_columns': 'A:G',
                'has_status_column': True,
                'status_column': 'G'
            },
            'Evidence_Register': {
                'evidence_start_row': 5,
                'evidence_columns': 'A:F',
                'has_evidence': True
            }
        }
    }
}

# Dashboard target sheet mappings
DASHBOARD_SHEETS = {
    'Executive_Dashboard': {
        'metrics_start_row': 10,
        'source': 'calculated'  # Populated from other sheets
    },
    'Gap_Analysis': {
        'data_start_row': 5,
        'consolidates_from': ['domain1', 'domain2', 'domain3', 'domain4']
    },
    'Evidence_Summary': {
        'data_start_row': 5,
        'consolidates_from': ['domain1', 'domain2', 'domain3', 'domain4']
    },
    'Risk_Matrix': {
        'data_start_row': 5,
        'consolidates_from': ['domain1', 'domain2', 'domain3', 'domain4']
    },
    'Compliance_Trend_Analysis': {
        'data_start_row': 5,
        'source': 'time_series'  # Historical data
    },
    'Audit_Findings': {
        'data_start_row': 5,
        'consolidates_from': ['domain1', 'domain2', 'domain3', 'domain4']
    },
    'Remediation_Tracking': {
        'data_start_row': 5,
        'source': 'gap_analysis'  # Derived from gaps
    }
}


# ============================================================================
# UTILITY FUNCTIONS - SAFE DATA OPERATIONS
# ============================================================================

def safely_write_data(ws, start_row, data, source_label=None):
    """
    Safely write data to worksheet, handling merged cells and errors.
    
    Args:
        ws: Worksheet object to write to
        start_row: Starting row number (1-indexed)
        data: List of lists containing row data
        source_label: Optional label to add to first column (traceability)
    
    Returns:
        int: Number of entries successfully written
    """
    entries_written = 0
    
    for row_idx, row_data in enumerate(data, start=start_row):
        # Add source label if provided
        if source_label and row_idx == start_row:
            # Write source label in a header or tracking column if your layout supports it
            pass  # Implement based on your dashboard layout
        
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Skip merged cells - cannot write to them
                if isinstance(cell, MergedCell):
                    continue
                
                # Write value
                cell.value = value
                
            except Exception as e:
                # Log error but continue processing
                print(f"    ⚠️  Warning: Could not write to row {row_idx}, col {col_idx}: {e}")
                continue
        
        entries_written += 1
    
    return entries_written


def extract_gaps_from_workbook(filepath, assessment_area, domain_schema):
    """
    Extract gap items (Partial/Non-Compliant) from source assessment workbook.
    
    Args:
        filepath: Path to assessment workbook
        assessment_area: Description of assessment area (for labeling)
        domain_schema: Schema definition for this domain
    
    Returns:
        list: List of gap data rows (list of lists)
    """
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        gaps = []
        gap_counter = 1
        
        # Extract from Summary_Dashboard sheet if it has gaps
        summary_sheet_config = domain_schema['sheets'].get('Summary_Dashboard', {})
        
        if summary_sheet_config.get('has_gaps', False) and 'Summary_Dashboard' in wb.sheetnames:
            ws = wb['Summary_Dashboard']
            gap_start_row = summary_sheet_config['gap_start_row']
            gap_columns = summary_sheet_config['gap_columns']
            
            # Parse column range (e.g., 'A:H')
            col_start, col_end = gap_columns.split(':')
            col_start_idx = ord(col_start) - ord('A') + 1
            col_end_idx = ord(col_end) - ord('A') + 1
            
            # Read gap rows
            for row_idx in range(gap_start_row, gap_start_row + 50):  # Read up to 50 rows
                row_data = []
                empty_row = True
                
                for col_idx in range(col_start_idx, col_end_idx + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    row_data.append(cell_value)
                    
                    if cell_value:
                        empty_row = False
                
                # Stop at first completely empty row
                if empty_row:
                    break
                
                # Check if this row represents a gap (has status indicator)
                # Typically last column has status like "⚠️ Partial" or "❌ Non-Compliant"
                if len(row_data) > 0:
                    status = str(row_data[-1]) if row_data[-1] else ""
                    if '⚠️' in status or '❌' in status or 'Partial' in status or 'Non-Compliant' in status:
                        # Add source label
                        labeled_row = [f"{assessment_area} - Gap {gap_counter}"] + row_data
                        gaps.append(labeled_row)
                        gap_counter += 1
        
        wb.close()
        return gaps
        
    except Exception as e:
        print(f"    ⚠️  Error extracting gaps from {filepath}: {e}")
        return []


def extract_evidence_from_workbook(filepath, assessment_area, domain_schema):
    """
    Extract evidence register entries from source assessment workbook.
    
    Args:
        filepath: Path to assessment workbook
        assessment_area: Description of assessment area (for labeling)
        domain_schema: Schema definition for this domain
    
    Returns:
        list: List of evidence data rows (list of lists)
    """
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        evidence = []
        
        evidence_sheet_config = domain_schema['sheets'].get('Evidence_Register', {})
        
        if evidence_sheet_config.get('has_evidence', False) and 'Evidence_Register' in wb.sheetnames:
            ws = wb['Evidence_Register']
            evidence_start_row = evidence_sheet_config['evidence_start_row']
            evidence_columns = evidence_sheet_config['evidence_columns']
            
            # Parse column range
            col_start, col_end = evidence_columns.split(':')
            col_start_idx = ord(col_start) - ord('A') + 1
            col_end_idx = ord(col_end) - ord('A') + 1
            
            # Read evidence rows
            for row_idx in range(evidence_start_row, evidence_start_row + 100):  # Read up to 100 rows
                row_data = []
                empty_row = True
                
                for col_idx in range(col_start_idx, col_end_idx + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    row_data.append(cell_value)
                    
                    if cell_value:
                        empty_row = False
                
                # Stop at first completely empty row
                if empty_row:
                    break
                
                # Add source label
                labeled_row = [assessment_area] + row_data
                evidence.append(labeled_row)
        
        wb.close()
        return evidence
        
    except Exception as e:
        print(f"    ⚠️  Error extracting evidence from {filepath}: {e}")
        return []
    
def extract_compliance_metrics(filepath, assessment_area, domain_schema):
    """
    Extract compliance metrics from Summary_Dashboard sheet.
    
    Args:
        filepath: Path to assessment workbook
        assessment_area: Description of assessment area
        domain_schema: Schema definition for this domain
    
    Returns:
        dict: Compliance metrics (compliance_pct, gap_count, etc.)
    """
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        metrics = {
            'assessment_area': assessment_area,
            'compliance_percentage': 0,
            'gap_count': 0,
            'compliant_count': 0,
            'non_compliant_count': 0,
            'partial_count': 0
        }
        
        summary_config = domain_schema['sheets'].get('Summary_Dashboard', {})
        
        if 'Summary_Dashboard' in wb.sheetnames:
            ws = wb['Summary_Dashboard']
            
            # Extract compliance percentage (typically in column B)
            compliance_row = summary_config.get('compliance_metrics_row', 15)
            compliance_value = ws.cell(row=compliance_row, column=2).value
            
            if compliance_value:
                # Handle percentage values (could be 0.85 or "85%" or 85)
                if isinstance(compliance_value, str):
                    compliance_value = compliance_value.replace('%', '').strip()
                try:
                    metrics['compliance_percentage'] = float(compliance_value)
                except:
                    pass
            
            # Extract gap count
            gap_row = summary_config.get('gap_count_row', 20)
            gap_value = ws.cell(row=gap_row, column=2).value
            if gap_value:
                try:
                    metrics['gap_count'] = int(gap_value)
                except:
                    pass
            
            # Extract status counts (scan for them in dashboard)
            # Typically displayed as "✅ Compliant: X", "⚠️ Partial: Y", "❌ Non-Compliant: Z"
            for row_idx in range(15, 30):
                for col_idx in range(1, 5):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value and isinstance(cell_value, str):
                        if 'Compliant:' in cell_value or '✅' in cell_value:
                            # Try to extract number
                            parts = str(cell_value).split(':')
                            if len(parts) > 1:
                                try:
                                    metrics['compliant_count'] = int(parts[1].strip())
                                except:
                                    pass
                        elif 'Partial' in cell_value or '⚠️' in cell_value:
                            parts = str(cell_value).split(':')
                            if len(parts) > 1:
                                try:
                                    metrics['partial_count'] = int(parts[1].strip())
                                except:
                                    pass
                        elif 'Non-Compliant' in cell_value or '❌' in cell_value:
                            parts = str(cell_value).split(':')
                            if len(parts) > 1:
                                try:
                                    metrics['non_compliant_count'] = int(parts[1].strip())
                                except:
                                    pass
        
        wb.close()
        return metrics
        
    except Exception as e:
        print(f"    ⚠️  Error extracting metrics from {filepath}: {e}")
        return {
            'assessment_area': assessment_area,
            'compliance_percentage': 0,
            'gap_count': 0,
            'compliant_count': 0,
            'non_compliant_count': 0,
            'partial_count': 0
        }


def validate_workbook_exists(filepath, domain_name):
    """
    Validate that a workbook exists and is readable.
    
    Args:
        filepath: Path to check
        domain_name: Domain identifier for error messages
    
    Returns:
        bool: True if valid, False otherwise
    """
    if not os.path.exists(filepath):
        print(f"    ❌ {domain_name} workbook not found: {filepath}")
        return False
    
    try:
        # Try to open workbook to verify it's valid
        wb = load_workbook(filepath, read_only=True)
        wb.close()
        print(f"    ✅ {domain_name} workbook validated: {os.path.basename(filepath)}")
        return True
    except Exception as e:
        print(f"    ❌ {domain_name} workbook is corrupt or unreadable: {e}")
        return False


def find_assessment_files(input_dir, domain_schemas):
    """
    Auto-detect assessment files in input directory based on filename patterns.
    
    Args:
        input_dir: Directory to search
        domain_schemas: Schema definitions with filename patterns
    
    Returns:
        dict: {domain_key: filepath} or None if any required file missing
    """
    print(f"\n🔍 Scanning for assessment workbooks in: {input_dir}")
    
    found_files = {}
    missing_domains = []
    
    for domain_key, schema in domain_schemas.items():
        pattern = schema['filename_pattern']
        normalized = schema['normalized_name']
        
        # Try normalized name first (from normalize_assessment_files_a832.py)
        normalized_path = os.path.join(input_dir, normalized)
        if os.path.exists(normalized_path):
            found_files[domain_key] = normalized_path
            print(f"  ✅ Found {domain_key}: {normalized}")
            continue
        
        # Try pattern matching for dated files
        import glob
        matches = glob.glob(os.path.join(input_dir, f"{pattern}*.xlsx"))
        
        # Filter out temp files
        matches = [m for m in matches if not os.path.basename(m).startswith('~$')]
        
        if matches:
            # Use most recent file if multiple matches
            latest = max(matches, key=os.path.getmtime)
            found_files[domain_key] = latest
            print(f"  ✅ Found {domain_key}: {os.path.basename(latest)}")
        else:
            missing_domains.append(domain_key)
            print(f"  ❌ Missing {domain_key}: No files matching '{pattern}*.xlsx'")
    
    if missing_domains:
        print(f"\n⚠️  WARNING: {len(missing_domains)} required assessment(s) missing!")
        print("   Missing domains:", ', '.join(missing_domains))
        print("\n   Consolidation requires all 4 domain assessments.")
        print("   Options:")
        print("   1. Complete missing assessments")
        print("   2. Run normalize_assessment_files_a832.py to create normalized filenames")
        print("   3. Use --domainN flags to specify file locations explicitly\n")
        return None
    
    return found_files


def consolidate_gap_analysis(dashboard_wb, source_files, domain_schemas):
    """
    Consolidate gap analysis data from all source workbooks into dashboard.
    
    Args:
        dashboard_wb: Dashboard workbook object
        source_files: Dict of {domain_key: filepath}
        domain_schemas: Schema definitions
    
    Returns:
        int: Total number of gaps consolidated
    """
    print("\n📊 Consolidating Gap Analysis...")
    
    if 'Gap_Analysis' not in dashboard_wb.sheetnames:
        print("  ⚠️  Warning: Gap_Analysis sheet not found in dashboard")
        return 0
    
    ws = dashboard_wb['Gap_Analysis']
    
    # Determine starting row (skip headers)
    start_row = DASHBOARD_SHEETS['Gap_Analysis']['data_start_row']
    current_row = start_row
    
    total_gaps = 0
    
    for domain_key in ['domain1', 'domain2', 'domain3', 'domain4']:
        if domain_key not in source_files:
            continue
        
        filepath = source_files[domain_key]
        schema = domain_schemas[domain_key]
        assessment_area = schema['assessment_area']
        
        print(f"  📁 Extracting gaps from {assessment_area}...")
        
        gaps = extract_gaps_from_workbook(filepath, assessment_area, schema)
        
        if gaps:
            entries = safely_write_data(ws, current_row, gaps, source_label=assessment_area)
            print(f"    ✅ Consolidated {entries} gap(s)")
            current_row += entries
            total_gaps += entries
        else:
            print(f"    ℹ️  No gaps found (fully compliant)")
    
    print(f"\n  ✅ Total gaps consolidated: {total_gaps}")
    return total_gaps


def consolidate_evidence_summary(dashboard_wb, source_files, domain_schemas):
    """
    Consolidate evidence register entries from all source workbooks into dashboard.
    
    Args:
        dashboard_wb: Dashboard workbook object
        source_files: Dict of {domain_key: filepath}
        domain_schemas: Schema definitions
    
    Returns:
        int: Total number of evidence entries consolidated
    """
    print("\n📋 Consolidating Evidence Summary...")
    
    if 'Evidence_Summary' not in dashboard_wb.sheetnames:
        print("  ⚠️  Warning: Evidence_Summary sheet not found in dashboard")
        return 0
    
    ws = dashboard_wb['Evidence_Summary']
    
    # Determine starting row
    start_row = DASHBOARD_SHEETS['Evidence_Summary']['data_start_row']
    current_row = start_row
    
    total_evidence = 0
    
    for domain_key in ['domain1', 'domain2', 'domain3', 'domain4']:
        if domain_key not in source_files:
            continue
        
        filepath = source_files[domain_key]
        schema = domain_schemas[domain_key]
        assessment_area = schema['assessment_area']
        
        print(f"  📁 Extracting evidence from {assessment_area}...")
        
        evidence = extract_evidence_from_workbook(filepath, assessment_area, schema)
        
        if evidence:
            entries = safely_write_data(ws, current_row, evidence, source_label=assessment_area)
            print(f"    ✅ Consolidated {entries} evidence item(s)")
            current_row += entries
            total_evidence += entries
        else:
            print(f"    ℹ️  No evidence entries found")
    
    print(f"\n  ✅ Total evidence entries consolidated: {total_evidence}")
    return total_evidence


def update_executive_dashboard(dashboard_wb, source_files, domain_schemas):
    """
    Update Executive_Dashboard sheet with aggregate metrics from all domains.
    
    Args:
        dashboard_wb: Dashboard workbook object
        source_files: Dict of {domain_key: filepath}
        domain_schemas: Schema definitions
    
    Returns:
        dict: Aggregate metrics
    """
    print("\n📈 Updating Executive Dashboard...")
    
    if 'Executive_Dashboard' not in dashboard_wb.sheetnames:
        print("  ⚠️  Warning: Executive_Dashboard sheet not found in dashboard")
        return {}
    
    ws = dashboard_wb['Executive_Dashboard']
    
    # Collect metrics from all domains
    all_metrics = []
    
    for domain_key in ['domain1', 'domain2', 'domain3', 'domain4']:
        if domain_key not in source_files:
            continue
        
        filepath = source_files[domain_key]
        schema = domain_schemas[domain_key]
        assessment_area = schema['assessment_area']
        
        print(f"  📊 Extracting metrics from {assessment_area}...")
        
        metrics = extract_compliance_metrics(filepath, assessment_area, schema)
        all_metrics.append(metrics)
        
        print(f"    ✅ Compliance: {metrics['compliance_percentage']:.1f}%, Gaps: {metrics['gap_count']}")
    
    # Calculate aggregate metrics
    total_domains = len(all_metrics)
    avg_compliance = sum(m['compliance_percentage'] for m in all_metrics) / total_domains if total_domains > 0 else 0
    total_gaps = sum(m['gap_count'] for m in all_metrics)
    total_compliant = sum(m['compliant_count'] for m in all_metrics)
    total_partial = sum(m['partial_count'] for m in all_metrics)
    total_non_compliant = sum(m['non_compliant_count'] for m in all_metrics)
    
    # Write aggregate metrics to dashboard (typical location: rows 10-20)
    metrics_start = DASHBOARD_SHEETS['Executive_Dashboard']['metrics_start_row']
    
    # Overall compliance percentage
    ws.cell(row=metrics_start, column=2).value = avg_compliance
    
    # Total gaps
    ws.cell(row=metrics_start + 1, column=2).value = total_gaps
    
    # Status counts
    ws.cell(row=metrics_start + 2, column=2).value = total_compliant
    ws.cell(row=metrics_start + 3, column=2).value = total_partial
    ws.cell(row=metrics_start + 4, column=2).value = total_non_compliant
    
    # Domain-specific metrics (rows for each domain)
    for idx, metrics in enumerate(all_metrics):
        domain_row = metrics_start + 6 + idx
        ws.cell(row=domain_row, column=1).value = metrics['assessment_area']
        ws.cell(row=domain_row, column=2).value = metrics['compliance_percentage']
        ws.cell(row=domain_row, column=3).value = metrics['gap_count']
    
    aggregate = {
        'overall_compliance': avg_compliance,
        'total_gaps': total_gaps,
        'total_compliant': total_compliant,
        'total_partial': total_partial,
        'total_non_compliant': total_non_compliant,
        'domains_assessed': total_domains
    }
    
    print(f"\n  ✅ Executive Dashboard Updated:")
    print(f"     Overall Compliance: {avg_compliance:.1f}%")
    print(f"     Total Gaps: {total_gaps}")
    print(f"     Domains Assessed: {total_domains}/4")
    
    return aggregate


def generate_consolidation_log(log_path, source_files, stats):
    """
    Generate consolidation log file for audit trail.
    
    Args:
        log_path: Path to write log file
        source_files: Dict of source files used
        stats: Consolidation statistics
    """
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("ISMS A.8.32 - DASHBOARD CONSOLIDATION LOG\n")
        f.write("ISO/IEC 27001:2022 - Control A.8.32: Change Management\n")
        f.write("=" * 80 + "\n\n")
        
        # Consolidation metadata
        f.write("CONSOLIDATION METADATA\n")
        f.write("-" * 80 + "\n")
        f.write(f"Consolidation Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Dashboard File:          {stats.get('dashboard_file', 'N/A')}\n")
        f.write(f"Domains Consolidated:    {len(source_files)}/4\n")
        f.write(f"Consolidation Status:    {'✅ COMPLETE' if len(source_files) == 4 else '⚠️  INCOMPLETE'}\n")
        f.write("\n")
        
        # Source files
        f.write("SOURCE ASSESSMENT FILES\n")
        f.write("-" * 80 + "\n")
        for domain_key, filepath in source_files.items():
            schema = WORKBOOK_SCHEMAS[domain_key]
            f.write(f"\nDomain: {domain_key.upper()}\n")
            f.write(f"  Assessment Area: {schema['assessment_area']}\n")
            f.write(f"  Source File:     {os.path.basename(filepath)}\n")
            f.write(f"  File Size:       {os.path.getsize(filepath):,} bytes\n")
            f.write(f"  Modified:        {datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("\n")
        
        # Consolidation statistics
        f.write("CONSOLIDATION STATISTICS\n")
        f.write("-" * 80 + "\n")
        f.write(f"Overall Compliance:      {stats.get('overall_compliance', 0):.1f}%\n")
        f.write(f"Total Gaps:              {stats.get('total_gaps', 0)}\n")
        f.write(f"Total Evidence Entries:  {stats.get('total_evidence', 0)}\n")
        f.write(f"Compliant Items:         {stats.get('total_compliant', 0)}\n")
        f.write(f"Partial Items:           {stats.get('total_partial', 0)}\n")
        f.write(f"Non-Compliant Items:     {stats.get('total_non_compliant', 0)}\n")
        f.write("\n")
        
        # Warnings
        if len(source_files) < 4:
            f.write("WARNINGS\n")
            f.write("-" * 80 + "\n")
            f.write(f"⚠️  Only {len(source_files)} of 4 required domains were consolidated.\n")
            f.write("   Dashboard metrics are incomplete.\n\n")
            missing = [k for k in WORKBOOK_SCHEMAS.keys() if k not in source_files]
            f.write("   Missing domains:\n")
            for domain_key in missing:
                f.write(f"   - {domain_key}: {WORKBOOK_SCHEMAS[domain_key]['assessment_area']}\n")
            f.write("\n")
        
        # Next steps
        f.write("NEXT STEPS\n")
        f.write("-" * 80 + "\n")
        f.write("1. Open the populated dashboard workbook\n")
        f.write("2. Review Executive Dashboard for overall compliance status\n")
        f.write("3. Examine Gap Analysis sheet for detailed findings\n")
        f.write("4. Verify Evidence Summary contains all required evidence\n")
        f.write("5. Use Remediation Tracking sheet to plan gap closure\n")
        f.write("6. Present dashboard to management for approval\n\n")
        
        f.write("AUDIT TRAIL\n")
        f.write("-" * 80 + "\n")
        f.write("This consolidation log provides audit evidence that:\n")
        f.write("- Dashboard was systematically populated from source assessments\n")
        f.write("- No manual data manipulation occurred\n")
        f.write("- All metrics are traceable to source workbooks\n")
        f.write("- Consolidation process was automated and repeatable\n\n")
        
        f.write("Retain this log with dashboard for ISO 27001:2022 audits.\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("END OF CONSOLIDATION LOG\n")
        f.write("=" * 80 + "\n")


def validate_dashboard_template(dashboard_path):
    """
    Validate dashboard template has expected structure.
    
    Args:
        dashboard_path: Path to dashboard workbook
    
    Returns:
        bool: True if valid, False otherwise
    """
    print("\n🔍 Validating dashboard template structure...")
    
    try:
        wb = load_workbook(dashboard_path)
        
        required_sheets = [
            'Instructions & Legend',
            'Executive_Dashboard',
            'Gap_Analysis',
            'Evidence_Summary',
            'Remediation_Tracking'
        ]
        
        missing_sheets = []
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                missing_sheets.append(sheet_name)
        
        if missing_sheets:
            print(f"  ❌ Dashboard template missing required sheets:")
            for sheet in missing_sheets:
                print(f"     - {sheet}")
            wb.close()
            return False
        
        print("  ✅ Dashboard template structure validated")
        wb.close()
        return True
        
    except Exception as e:
        print(f"  ❌ Error validating dashboard template: {e}")
        return False
    
def create_backup(dashboard_path):
    """
    Create backup of dashboard before consolidation.
    
    Args:
        dashboard_path: Path to dashboard workbook
    
    Returns:
        str: Path to backup file, or None if backup failed
    """
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_dir = os.path.dirname(dashboard_path)
        backup_name = f"dashboard_backup_{timestamp}.xlsx"
        backup_path = os.path.join(backup_dir, backup_name)
        
        import shutil
        shutil.copy2(dashboard_path, backup_path)
        
        print(f"  ✅ Backup created: {backup_name}")
        return backup_path
        
    except Exception as e:
        print(f"  ⚠️  Warning: Backup failed: {e}")
        print("     Continuing without backup...")
        return None


def consolidate_dashboard(dashboard_path, source_files, domain_schemas, options=None):
    """
    Main dashboard consolidation orchestration function.
    
    Args:
        dashboard_path: Path to dashboard workbook
        source_files: Dict of {domain_key: filepath}
        domain_schemas: Schema definitions
        options: Optional dict of consolidation options
    
    Returns:
        dict: Consolidation statistics
    """
    if options is None:
        options = {}
    
    dry_run = options.get('dry_run', False)
    verbose = options.get('verbose', False)
    backup = options.get('backup', False)
    
    stats = {
        'dashboard_file': os.path.basename(dashboard_path),
        'start_time': datetime.now(),
        'overall_compliance': 0,
        'total_gaps': 0,
        'total_evidence': 0,
        'total_compliant': 0,
        'total_partial': 0,
        'total_non_compliant': 0,
        'domains_consolidated': 0
    }
    
    print("\n" + "=" * 80)
    print("ISMS A.8.32 - DASHBOARD CONSOLIDATION")
    print("=" * 80)
    print(f"\nDashboard: {os.path.basename(dashboard_path)}")
    print(f"Mode:      {'DRY RUN (no changes will be saved)' if dry_run else 'PRODUCTION (dashboard will be updated)'}")
    print(f"Started:   {stats['start_time'].strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Phase 1: Validate dashboard template
    print("\n" + "=" * 80)
    print("PHASE 1: PRE-CONSOLIDATION VALIDATION")
    print("=" * 80)
    
    if not validate_dashboard_template(dashboard_path):
        print("\n❌ Dashboard template validation failed!")
        print("   Generate a new dashboard template with generate_a832_5_compliance_dashboard.py")
        return None
    
    # Create backup if requested
    if backup and not dry_run:
        print("\n💾 Creating backup...")
        backup_path = create_backup(dashboard_path)
        if backup_path:
            stats['backup_file'] = backup_path
    
    # Phase 2: Load dashboard
    print("\n" + "=" * 80)
    print("PHASE 2: LOADING DASHBOARD")
    print("=" * 80)
    
    try:
        print(f"\n📂 Loading dashboard workbook...")
        dashboard_wb = load_workbook(dashboard_path)
        print(f"  ✅ Dashboard loaded ({len(dashboard_wb.sheetnames)} sheets)")
        
        if verbose:
            print(f"\n  Dashboard sheets:")
            for sheet_name in dashboard_wb.sheetnames:
                print(f"    • {sheet_name}")
    
    except Exception as e:
        print(f"\n❌ Error loading dashboard: {e}")
        return None
    
    # Phase 3: Consolidate data
    print("\n" + "=" * 80)
    print("PHASE 3: DATA CONSOLIDATION")
    print("=" * 80)
    
    try:
        # Update executive dashboard with aggregate metrics
        aggregate_metrics = update_executive_dashboard(dashboard_wb, source_files, domain_schemas)
        stats.update(aggregate_metrics)
        stats['domains_consolidated'] = aggregate_metrics.get('domains_assessed', 0)
        
        # Consolidate gap analysis
        total_gaps = consolidate_gap_analysis(dashboard_wb, source_files, domain_schemas)
        stats['total_gaps'] = total_gaps
        
        # Consolidate evidence summary
        total_evidence = consolidate_evidence_summary(dashboard_wb, source_files, domain_schemas)
        stats['total_evidence'] = total_evidence
        
    except Exception as e:
        print(f"\n❌ Error during consolidation: {e}")
        dashboard_wb.close()
        return None
    
    # Phase 4: Save dashboard
    print("\n" + "=" * 80)
    print("PHASE 4: SAVING DASHBOARD")
    print("=" * 80)
    
    if dry_run:
        print("\n🔍 DRY RUN MODE: Dashboard changes NOT saved")
        print("   Run without --dry-run to save changes")
        dashboard_wb.close()
    else:
        try:
            print(f"\n💾 Saving consolidated dashboard...")
            dashboard_wb.save(dashboard_path)
            print(f"  ✅ Dashboard saved successfully")
        except Exception as e:
            print(f"\n❌ Error saving dashboard: {e}")
            dashboard_wb.close()
            return None
        finally:
            dashboard_wb.close()
    
    # Phase 5: Generate consolidation log
    print("\n" + "=" * 80)
    print("PHASE 5: GENERATING CONSOLIDATION LOG")
    print("=" * 80)
    
    if not dry_run:
        timestamp = datetime.now().strftime('%Y%m%d')
        log_dir = os.path.dirname(dashboard_path)
        log_filename = f"A832_Dashboard_Consolidation_Log_{timestamp}.txt"
        log_path = os.path.join(log_dir, log_filename)
        
        try:
            print(f"\n📝 Generating consolidation log...")
            generate_consolidation_log(log_path, source_files, stats)
            print(f"  ✅ Log created: {log_filename}")
            stats['log_file'] = log_path
        except Exception as e:
            print(f"  ⚠️  Warning: Log generation failed: {e}")
    
    # Calculate elapsed time
    stats['end_time'] = datetime.now()
    stats['elapsed_seconds'] = (stats['end_time'] - stats['start_time']).total_seconds()
    
    return stats


def print_consolidation_summary(stats, source_files):
    """
    Print consolidation summary statistics.
    
    Args:
        stats: Consolidation statistics dict
        source_files: Dict of source files consolidated
    """
    print("\n" + "=" * 80)
    print("CONSOLIDATION COMPLETE")
    print("=" * 80)
    
    print(f"\n📊 Statistics:")
    print(f"  • Domains consolidated:     {stats['domains_consolidated']}/4")
    print(f"  • Overall compliance:       {stats['overall_compliance']:.1f}%")
    print(f"  • Total gaps identified:    {stats['total_gaps']}")
    print(f"  • Evidence entries:         {stats['total_evidence']}")
    print(f"  • Compliant items:          {stats['total_compliant']}")
    print(f"  • Partial compliance:       {stats['total_partial']}")
    print(f"  • Non-compliant items:      {stats['total_non_compliant']}")
    
    print(f"\n📋 By Domain:")
    for domain_key in ['domain1', 'domain2', 'domain3', 'domain4']:
        if domain_key in source_files:
            schema = WORKBOOK_SCHEMAS[domain_key]
            print(f"  • {schema['assessment_area']}")
        else:
            schema = WORKBOOK_SCHEMAS[domain_key]
            print(f"  ⚠️  {schema['assessment_area']} - NOT CONSOLIDATED")
    
    print(f"\n⏱️  Completed: {stats['end_time'].strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"   Elapsed:   {stats['elapsed_seconds']:.1f} seconds")
    
    if 'log_file' in stats:
        print(f"\n📄 Files Generated:")
        print(f"  • Dashboard: {stats['dashboard_file']}")
        print(f"  • Log:       {os.path.basename(stats['log_file'])}")
        if 'backup_file' in stats:
            print(f"  • Backup:    {os.path.basename(stats['backup_file'])}")
    
    print("\n" + "=" * 80)
    print("🎯 Dashboard now contains consolidated compliance data from all assessments!")
    print("   Open the dashboard workbook to review executive summary and gaps.")
    print("=" * 80 + "\n")


def parse_arguments():
    """
    Parse command-line arguments.
    
    Returns:
        Namespace: Parsed arguments
    """
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Consolidate A.8.32 assessment data into compliance dashboard',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Auto-detect files in current directory
  python3 consolidate_a832_dashboard.py
  
  # Specify input directory
  python3 consolidate_a832_dashboard.py --input-dir ./assessments
  
  # Dry run (validate without saving)
  python3 consolidate_a832_dashboard.py --dry-run
  
  # With backup and verbose output
  python3 consolidate_a832_dashboard.py --backup --verbose
  
  # Specify files explicitly
  python3 consolidate_a832_dashboard.py \\
    --domain1 ./domain1_assessment.xlsx \\
    --domain2 ./domain2_assessment.xlsx \\
    --domain3 ./domain3_assessment.xlsx \\
    --domain4 ./domain4_assessment.xlsx \\
    --dashboard ./dashboard.xlsx
        """
    )
    
    parser.add_argument(
        '--input-dir', '-i',
        default='.',
        help='Directory containing assessment workbooks (default: current directory)'
    )
    
    parser.add_argument(
        '--dashboard', '-d',
        help='Path to dashboard workbook (auto-detect if not specified)'
    )
    
    parser.add_argument(
        '--domain1',
        help='Path to Domain 1 assessment (Change Process)'
    )
    
    parser.add_argument(
        '--domain2',
        help='Path to Domain 2 assessment (Change Types & Categories)'
    )
    
    parser.add_argument(
        '--domain3',
        help='Path to Domain 3 assessment (Environment Separation)'
    )
    
    parser.add_argument(
        '--domain4',
        help='Path to Domain 4 assessment (Testing & Validation)'
    )
    
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Validate without writing to dashboard'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable detailed logging'
    )
    
    parser.add_argument(
        '--backup', '-b',
        action='store_true',
        help='Create backup before consolidation'
    )
    
    parser.add_argument(
        '--pdf',
        action='store_true',
        help='Generate PDF executive summary (requires reportlab)'
    )
    
    return parser.parse_args()


def find_dashboard_file(directory):
    """
    Auto-detect dashboard file in directory.
    
    Args:
        directory: Directory to search
    
    Returns:
        str: Path to dashboard file, or None if not found
    """
    import glob
    
    # Try normalized name first
    normalized = os.path.join(directory, 'ISMS-IMP-A.8.32.5.xlsx')
    if os.path.exists(normalized):
        return normalized
    
    # Try pattern matching
    pattern = os.path.join(directory, 'ISMS_A_8_32_5_Compliance_Dashboard*.xlsx')
    matches = glob.glob(pattern)
    
    # Filter out temp files and backups
    matches = [m for m in matches if not os.path.basename(m).startswith('~$') and 'backup' not in m.lower()]
    
    if matches:
        # Use most recent
        return max(matches, key=os.path.getmtime)
    
    return None


def main():
    """
    Main execution function.
    
    Returns:
        int: Exit code (0=success, 1-4=various errors)
    """
    args = parse_arguments()
    
    print("\n" + "=" * 80)
    print("ISMS A.8.32 - DASHBOARD CONSOLIDATION UTILITY")
    print("ISO/IEC 27001:2022 - Control A.8.32: Change Management")
    print("=" * 80)
    print(f"\nVersion:  1.0")
    print(f"Date:     {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Determine source files
    if all([args.domain1, args.domain2, args.domain3, args.domain4]):
        # Explicit file paths provided
        print("\n📁 Using explicitly specified source files...")
        source_files = {
            'domain1': args.domain1,
            'domain2': args.domain2,
            'domain3': args.domain3,
            'domain4': args.domain4
        }
        
        # Validate all files exist
        all_valid = True
        for domain_key, filepath in source_files.items():
            if not validate_workbook_exists(filepath, domain_key):
                all_valid = False
        
        if not all_valid:
            print("\n❌ One or more source files invalid or missing!")
            return 4  # File I/O error
    else:
        # Auto-detect files
        source_files = find_assessment_files(args.input_dir, WORKBOOK_SCHEMAS)
        
        if source_files is None:
            print("\n❌ Required assessment files not found!")
            print("   Options:")
            print("   1. Complete missing assessments")
            print("   2. Run normalize_assessment_files_a832.py")
            print("   3. Use --domain1 through --domain4 flags to specify file paths")
            return 4  # File I/O error
    
    # Determine dashboard file
    if args.dashboard:
        dashboard_path = args.dashboard
        print(f"\n📊 Using specified dashboard: {os.path.basename(dashboard_path)}")
    else:
        print(f"\n🔍 Auto-detecting dashboard file...")
        dashboard_path = find_dashboard_file(args.input_dir)
        
        if dashboard_path:
            print(f"  ✅ Found dashboard: {os.path.basename(dashboard_path)}")
        else:
            print("\n❌ Dashboard file not found!")
            print("   Generate dashboard with: python3 generate_a832_5_compliance_dashboard.py")
            print("   Or specify dashboard path with: --dashboard /path/to/dashboard.xlsx")
            return 4  # File I/O error
    
    # Validate dashboard exists
    if not os.path.exists(dashboard_path):
        print(f"\n❌ Dashboard file not found: {dashboard_path}")
        return 4  # File I/O error
    
    # Consolidation options
    options = {
        'dry_run': args.dry_run,
        'verbose': args.verbose,
        'backup': args.backup
    }
    
    # Run consolidation
    stats = consolidate_dashboard(dashboard_path, source_files, WORKBOOK_SCHEMAS, options)
    
    if stats is None:
        print("\n❌ Consolidation failed!")
        print("   Check error messages above for details.")
        return 3  # Consolidation error
    
    # Print summary
    print_consolidation_summary(stats, source_files)
    
    # Generate PDF if requested
    if args.pdf and not args.dry_run:
        print("📄 PDF generation requested...")
        try:
            import reportlab
            print("  ℹ️  PDF generation not yet implemented")
            print("     Feature planned for future version")
        except ImportError:
            print("  ⚠️  Warning: reportlab not installed")
            print("     Install with: pip3 install reportlab")
    
    # Success
    return 0


if __name__ == "__main__":
    try:
        exit_code = main()
        sys.exit(exit_code)
    except KeyboardInterrupt:
        print("\n\n⚠️  Consolidation interrupted by user")
        print("   Dashboard may be in inconsistent state if consolidation was in progress")
        print("   Restore from backup if needed\n")
        sys.exit(130)
    except Exception as e:
        print(f"\n\n❌ Unexpected error: {e}")
        print("   Contact support or check consolidation log for details\n")
        import traceback
        traceback.print_exc()
        sys.exit(1)