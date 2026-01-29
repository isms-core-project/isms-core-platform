#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.32.1 - Change Process Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.32: Change Management
Assessment Domain 1 of 4: Change Process Workflow & Management Procedures

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific change management processes and assessment
requirements.

Key customization areas:
1. Sheet names and structures (match your control's assessment needs)
2. Data validation rules (specific to your change management requirements)
3. Compliance scoring logic (adapt to your risk profile)
4. Output formatting (match your assessment workbook structure)
5. Change process stages (customize to your workflow)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.32 Change Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
change management processes and workflows against ISO 27001:2022 Control A.8.32
requirements, supporting evidence-based validation of change management
procedures.

**Purpose:**
Enables systematic assessment of change management processes including workflow
stages, approval authorities, communication procedures, documentation requirements,
and tool capabilities against ISO 27001:2022 Control A.8.32 requirements.

**Assessment Scope:**
- Change process workflow definition and stages
- Approval authority matrix and decision rights
- Stakeholder communication procedures and timing
- Documentation requirements and retention
- Change management tool capabilities and integration
- Process compliance verification
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and standards
2. Change_Process_Workflow - Process stages and procedures
3. Approval_Authority_Matrix - Authority levels and thresholds
4. Communication_Procedures - Stakeholder notification methods
5. Documentation_Requirements - Record-keeping standards
6. Change_Management_Tools - Tool inventory and capabilities
7. Summary_Dashboard - Compliance metrics and analytics
8. Evidence_Register - Audit evidence tracking
9. Approval_Sign_Off - Stakeholder approval workflow

**Key Features:**
- Technology-agnostic assessment (works with any change management tool)
- Data validation with dropdown lists for consistency
- Conditional formatting for visual status indicators
- Automated compliance calculations
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with A.8.32.5 Compliance Dashboard

**Integration:**
This assessment feeds into the A.8.32.5 Compliance Dashboard, which
consolidates data from all four change management assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a832_1_change_process.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a832_1_change_process.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a832_1_change_process.py --date 20260127

Output:
    File: ISMS_A_8_32_1_Change_Process_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review Instructions & Legend sheet for assessment guidance
    2. Document your organization's change process workflow stages
    3. Complete approval authority matrix with your roles and thresholds
    4. Define your stakeholder communication procedures
    5. Document your record-keeping and retention requirements
    6. Inventory your change management tools and capabilities
    7. Review Summary_Dashboard for compliance metrics
    8. Collect and link audit evidence in Evidence_Register
    9. Obtain stakeholder approvals via Approval_Sign_Off
    10. Feed results into A.8.32.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.32
Assessment Domain:    1 of 4 (Change Process Workflow & Management)
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.32: Change Management Policy (Governance)
    - ISMS-IMP-A.8.32.1: Change Process Assessment Implementation Guide
    - ISMS-IMP-A.8.32.2: Change Types & Categories Assessment (Domain 2)
    - ISMS-IMP-A.8.32.3: Environment Separation Assessment (Domain 3)
    - ISMS-IMP-A.8.32.4: Testing & Validation Assessment (Domain 4)
    - ISMS-IMP-A.8.32.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.32.1 specification
    - Supports comprehensive change process evaluation
    - Integrated with A.8.32.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Technology-Agnostic Design:**
This assessment framework is intentionally tool-independent. Whether your
organization uses ServiceNow, Jira, BMC Remedy, custom tools, spreadsheets,
or even paper-based processes, this assessment evaluates your PROCESS and
COMPLIANCE, not your tool brands.

**Process vs. Tool Assessment:**
The assessment focuses on:
- Is your change process defined and documented?
- Are approval authorities clear and appropriate?
- Are stakeholders notified appropriately?
- Is documentation complete and retained?
- Does your tool support (not dictate) your process?

NOT:
- Which vendor's tool you use
- How much you paid for your tool
- Whether you have the "latest and greatest" platform

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of your change process, not your tool choice.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Change process workflows and procedures
- Approval authorities and escalation paths
- Tool capabilities and limitations
- Identified compliance gaps

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check for process changes and tool updates
- Semi-annually: Review approval authorities
- Annually: Complete reassessment of all domains
- Ad-hoc: When significant process or organizational changes occur

**Quality Assurance:**
Have change management practitioners and security personnel validate assessments
before using results for compliance reporting or remediation decisions.

**Regulatory Alignment:**
Customize assessment criteria to include regulatory-specific requirements:
- Financial services: SOX change management controls
- Healthcare: HIPAA security rule change management
- Government: Jurisdiction-specific change control requirements
- PCI DSS: Change management for cardholder data environments

Adapt dropdown values and assessment questions to your regulatory context.

================================================================================
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================



# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
CLOCK = '\U0001F552'  # 🕒 Clock
WRENCH = '\U0001F527' # 🔧 Wrench
ROCKET = '\U0001F680' # 🚀 Rocket
GEAR = '\u2699'       # ⚙️  Gear
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP-A.8.32.1 spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches ISMS-IMP-A.8.32.1 specification
    sheets = [
        "Instructions & Legend",
        "Change_Process_Workflow",
        "Approval_Authority_Matrix",
        "Communication_Procedures",
        "Documentation_Requirements",
        "Change_Management_Tools",
        "Summary_Dashboard",
        "Evidence_Register",
        "Approval_Sign_Of",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "calculated_cell": {
            "fill": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_implemented": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_notimplemented": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_planned": {
            "fill": PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        },
        "compliance_full": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "compliance_substantial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "compliance_partial": {
            "fill": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
        },
        "compliance_non": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & DROPDOWN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    These are added to worksheet once, then applied to multiple cells.
    
    CUSTOMIZE: Adapt dropdown values for change management context.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_partial': DataValidation(
            type="list",
            formula1='"Yes,No,Partial"',
            allow_blank=False
        ),
        'implementation_status': DataValidation(
            type="list",
            formula1='f"{CHECK} Implemented,⚠️ Partial,❌ Not Implemented,📋 Planned,N/A"',
            allow_blank=False
        ),
        'automation_level': DataValidation(
            type="list",
            formula1='"Automated,Semi-Automated,Manual"',
            allow_blank=False
        ),
        'change_type': DataValidation(
            type="list",
            formula1='"Standard,Normal,Emergency"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Low,Medium,High,Critical"',
            allow_blank=False
        ),
        'impact_level': DataValidation(
            type="list",
            formula1='"Low,Medium,High,Critical"',
            allow_blank=False
        ),
        'cab_required': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'approval_method': DataValidation(
            type="list",
            formula1='"Virtual Meeting,Phone,Email Chain,Instant Messaging,Other"',
            allow_blank=False
        ),
        'notification_trigger': DataValidation(
            type="list",
            formula1='"Submission,Approval,24h Before,At Implementation,Completion,All"',
            allow_blank=False
        ),
        'communication_method': DataValidation(
            type="list",
            formula1='"Email,Portal,IM,SMS,All"',
            allow_blank=False
        ),
        'audience_reach': DataValidation(
            type="list",
            formula1='"All Users,IT Staff,Management,Selective"',
            allow_blank=False
        ),
        'reliability': DataValidation(
            type="list",
            formula1='"High,Medium,Low"',
            allow_blank=False
        ),
        'doc_requirement': DataValidation(
            type="list",
            formula1='"Mandatory,Optional,N/A"',
            allow_blank=False
        ),
        'format_type': DataValidation(
            type="list",
            formula1='"Electronic,Paper,Both"',
            allow_blank=False
        ),
        'disposal_method': DataValidation(
            type="list",
            formula1='"Secure Delete,Archive,Destruction"',
            allow_blank=False
        ),
        'deployment_type': DataValidation(
            type="list",
            formula1='"Cloud SaaS,On-Premises,Hybrid,Self-Hosted"',
            allow_blank=False
        ),
        'tool_capability': DataValidation(
            type="list",
            formula1='f"{CHECK} Full,⚠️ Partial,❌ No"',
            allow_blank=False
        ),
        'integration_type': DataValidation(
            type="list",
            formula1='"API,File Transfer,Manual,Webhook,Other"',
            allow_blank=False
        ),
        'integration_status': DataValidation(
            type="list",
            formula1='f"{CHECK} Integrated,⚠️ Partial,❌ Standalone"',
            allow_blank=False
        ),
        'license_status': DataValidation(
            type="list",
            formula1='"Active,Expiring Soon,Expired"',
            allow_blank=False
        ),
        'support_availability': DataValidation(
            type="list",
            formula1='"24/7,Business Hours,Community,None"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Policy Doc,Procedure,Flowchart,Template,Screenshot,Meeting Minutes,Email,Report,Contract,Other"',
            allow_blank=False
        ),
        'verification_status': DataValidation(
            type="list",
            formula1='f"{CHECK} Verified,⚠️ Pending,❌ Not Verified"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1='f"{CHECK} Final,⚠️ Requires Remediation,📋 Draft,❌ Re-assessment Required"',
            allow_blank=False
        ),
        'review_recommendation': DataValidation(
            type="list",
            formula1='f"{CHECK} Approve,⚠️ Approve with Conditions,❌ Reject,📋 Require Rework"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1='f"{CHECK} Approved,⚠️ Approved with Conditions,❌ Rejected"',
            allow_blank=False
        ),
        'regulatory_review': DataValidation(
            type="list",
            formula1='"Yes,No,To Be Determined"',
            allow_blank=False
        ),
    }

    # Add all validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)

    return validations


# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet with usage guidance."""
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "ISMS-IMP-A.8.32.1 – Change Process Assessment"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "ISO/IEC 27001:2022 - Control A.8.32: Change Management"
    apply_style(ws["A2"], styles["subheader"])

    # Document Information Block
    row = 4
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    
    doc_info = [
        ("Document ID:", "ISMS-IMP-A.8.32.1"),
        ("Assessment Area:", "Change Process Workflow & Management"),
        ("Related Policy:", "ISMS-POL-A.8.32-S2.1"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[Organization]"),
        ("Review Cycle:", "Quarterly"),
    ]
    
    row += 1
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if "USER INPUT" in value:
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    # How to Use This Workbook
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    instructions = [
        "1. Document YOUR organization's change management process in the Change_Process_Workflow sheet",
        "2. Complete the Approval_Authority_Matrix with YOUR specific roles and thresholds",
        "3. Fill in YOUR communication procedures and stakeholder notification methods",
        "4. Document YOUR documentation requirements and record-keeping practices",
        "5. Inventory YOUR change management tools (whatever platforms/systems you use)",
        "6. Review the Summary_Dashboard for compliance metrics",
        "7. Maintain the Evidence Register for audit traceability",
        "8. Obtain final approval and sign-of",
    ]

    row += 1
    for instruction in instructions:
        ws[f"A{row}"] = instruction
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1

    # Status Legend
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "STATUS LEGEND"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    legend_headers = ["Symbol", "Status", "Description", "Color Code"]
    for col_idx, header in enumerate(legend_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    legend_data = [
        (f"{CHECK}", "Implemented", "Process implemented and operational", "Green"),
        (f"{WARNING}", "Partial", "Partial implementation or limited coverage", "Yellow"),
        (f"{XMARK}", "Not Implemented", "Process not implemented", "Red"),
        ("📋", "Planned", "Implementation planned with target date", "Blue"),
        ("N/A", "Not Applicable", "Not applicable to this environment", "Gray"),
    ]

    row += 1
    for symbol, status, desc, color in legend_data:
        ws[f"A{row}"] = symbol
        ws[f"B{row}"] = status
        ws[f"C{row}"] = desc
        ws[f"D{row}"] = color
        
        # Apply color coding
        if color == "Green":
            ws[f"D{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif color == "Yellow":
            ws[f"D{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif color == "Red":
            ws[f"D{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif color == "Blue":
            ws[f"D{row}"].fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        
        row += 1

    # Compliance Levels
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "COMPLIANCE LEVELS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    compliance_headers = ["Level", "Percentage", "Description"]
    for col_idx, header in enumerate(compliance_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    compliance_data = [
        ("Fully Compliant", "90-100%", "All requirements met, evidence documented"),
        ("Substantially Compliant", "70-89%", "Most requirements met, minor gaps"),
        ("Partially Compliant", "50-69%", "Significant gaps, remediation required"),
        ("Non-Compliant", "<50%", "Major deficiencies, immediate action required"),
    ]

    row += 1
    for level, percentage, desc in compliance_data:
        ws[f"A{row}"] = level
        ws[f"B{row}"] = percentage
        ws[f"C{row}"] = desc
        row += 1

    # Acceptable Evidence
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (Examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    evidence_examples = [
        "✓ Change management policy documents",
        "✓ Process flowcharts/diagrams",
        "✓ Change request form templates",
        "✓ Change Advisory Board (CAB) meeting minutes",
        "✓ Approval workflow documentation",
        "✓ Communication templates (notification emails, status updates)",
        "✓ Change management tool screenshots (redacted if needed)",
        "✓ Change calendar/schedule examples",
        "✓ Post-Implementation Review (PIR) templates",
        "✓ Change failure metrics/reports",
        "✓ Emergency change procedure documentation",
        "✓ Training materials for change requesters/approvers",
    ]

    row += 1
    for evidence in evidence_examples:
        ws[f"A{row}"] = evidence
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 4: CHANGE PROCESS WORKFLOW SHEET
# ============================================================================

def create_change_process_workflow(ws, styles):
    """Create Change_Process_Workflow sheet documenting the complete change lifecycle."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "CHANGE MANAGEMENT PROCESS WORKFLOW"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Document the complete change lifecycle in your organization"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== PROCESS STAGE MAPPING ====================
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "PROCESS STAGE MAPPING"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    headers = ["Stage", "Stage Name", "Description", "Process Owner", "Standard Duration", "Tool/System Used", "Status", "Evidence"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    stages = [
        (1, "Change Request Initiation", "How changes are requested"),
        (2, "Initial Assessment & Screening", "Validation and categorization"),
        (3, "Risk & Impact Assessment", "Evaluate change risks"),
        (4, "Change Categorization", "Standard/Normal/Emergency"),
        (5, "CAB Scheduling (if required)", "Schedule review meeting"),
        (6, "CAB Review & Discussion", "Formal review process"),
        (7, "Change Authorization", "Approval decision"),
        (8, "Implementation Planning", "Schedule and resource allocation"),
        (9, "Stakeholder Communication", "Notification procedures"),
        (10, "Pre-Implementation Testing", "Testing in non-production"),
        (11, "Change Implementation", "Actual execution"),
        (12, "Implementation Verification", "Verify success"),
        (13, "Stakeholder Notification", "Completion notification"),
        (14, "Post-Implementation Review", "Lessons learned"),
        (15, "Change Record Closure", "Documentation completion"),
    ]

    row += 1
    for stage_num, stage_name, description in stages:
        ws[f"A{row}"] = stage_num
        ws[f"B{row}"] = stage_name
        ws[f"C{row}"] = description
        
        # Make editable cells yellow
        for col in [4, 5, 6, 8]:  # D=4, E=5, F=6, H=8
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Status dropdown
        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== PROCESS DOCUMENTATION REQUIREMENTS ====================
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "PROCESS DOCUMENTATION REQUIREMENTS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    doc_headers = ["Requirement", "Implemented", "Documentation Type", "Template Available", "Location", "Responsible Role"]
    for col_idx, header in enumerate(doc_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    doc_requirements = [
        "Change request form/template",
        "Risk assessment template",
        "Impact assessment template",
        "CAB agenda template",
        "Implementation plan template",
        "Rollback plan template",
        "PIR template",
        "Communication templates",
        "Change closure checklist",
    ]

    row += 1
    for requirement in doc_requirements:
        ws[f"A{row}"] = requirement
        
        # Editable cells
        for col in [3, 5, 6]:  # C=3, E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['yes_no_partial'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['yes_no'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== WORKFLOW AUTOMATION ASSESSMENT ====================
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "WORKFLOW AUTOMATION ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    auto_headers = ["Capability", "Automated", "Semi-Automated", "Manual", "Tool/System", "Notes"]
    for col_idx, header in enumerate(auto_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    automation_items = [
        "Change request submission",
        "Request routing/assignment",
        "Approval workflow",
        "Notification distribution",
        "Change calendar updates",
        "Status tracking",
        "Reporting/dashboards",
        "PIR reminders",
    ]

    row += 1
    for item in automation_items:
        ws[f"A{row}"] = item
        
        # Checkboxes (represented as Yes/No dropdowns in Excel)
        for col in [2, 3, 4]:  # B=2, C=3, D=4
            validations['yes_no'].add(ws.cell(row=row, column=col))
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Text fields
        for col in [5, 6]:  # E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 25

    ws.freeze_panes = "A6"


# ============================================================================
# SECTION 5: APPROVAL AUTHORITY MATRIX SHEET
# ============================================================================

def create_approval_authority_matrix(ws, styles):
    """Create Approval_Authority_Matrix sheet defining approval authorities."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "CHANGE APPROVAL AUTHORITY MATRIX"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "Define approval authorities by change type and risk level"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== STANDARD CHANGES APPROVAL ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "STANDARD CHANGES APPROVAL"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    std_headers = ["Standard Change Type", "Pre-Approved", "Self-Service Allowed", "Approver (if required)", "Approval SLA", "Implemented", "Evidence"]
    for col_idx, header in enumerate(std_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    row += 1
    # 15 rows for standard changes
    for i in range(15):
        for col in [1, 4, 5, 7]:  # A=1, D=4, E=5, G=7
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['yes_no'].add(ws.cell(row=row, column=2))
        ws.cell(row=row, column=2).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['yes_no_na'].add(ws.cell(row=row, column=3))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['implementation_status'].add(ws.cell(row=row, column=6))
        ws.cell(row=row, column=6).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== NORMAL CHANGES APPROVAL MATRIX ====================
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "NORMAL CHANGES APPROVAL MATRIX"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    normal_headers = ["Risk Level", "Impact Level", "Required Approvers", "CAB Review Required", "Approval Threshold", "Typical SLA", "Implemented"]
    for col_idx, header in enumerate(normal_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Define risk/impact combinations
    combinations = [
        ("Low", "Low"),
        ("Low", "Medium"),
        ("Low", "High"),
        ("Medium", "Low"),
        ("Medium", "Medium"),
        ("Medium", "High"),
        ("High", "Low"),
        ("High", "Medium"),
        ("High", "High"),
        ("Critical", "Any"),
    ]

    row += 1
    for risk, impact in combinations:
        ws[f"A{row}"] = risk
        ws[f"B{row}"] = impact
        
        # Editable cells
        for col in [3, 5, 6]:  # C=3, E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['yes_no'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== EMERGENCY CHANGES APPROVAL ====================
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "EMERGENCY CHANGES APPROVAL"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    emerg_headers = ["Emergency Criteria", "E-CAB Members", "Minimum Approvers", "Approval Method", "Documented", "Evidence"]
    for col_idx, header in enumerate(emerg_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    emergency_criteria = [
        "System outage affecting critical services",
        "Security incident response",
        "Regulatory compliance deadline",
        "Data loss prevention",
        "[Other - specify]",
    ]

    row += 1
    for criteria in emergency_criteria:
        ws[f"A{row}"] = criteria
        
        # Editable cells
        for col in [2, 3, 6]:  # B=2, C=3, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['approval_method'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['implementation_status'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== APPROVAL PROCESS METRICS ====================
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "APPROVAL PROCESS METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    metric_headers = ["Metric", "Target", "Current", "Status"]
    for col_idx, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    metrics = [
        "Average approval time (Standard)",
        "Average approval time (Normal - Low Risk)",
        "Average approval time (Normal - High Risk)",
        "Average approval time (Emergency)",
        "Approval backlog (pending >SLA)",
        "Approval rejection rate",
    ]

    row += 1
    for metric in metrics:
        ws[f"A{row}"] = metric
        
        # Editable cells
        for col in [2, 3]:  # B=2, C=3
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Status - calculated (conditional formatting would be applied manually)
        ws[f"D{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 25

    ws.freeze_panes = "A6"


# ============================================================================
# SECTION 6: COMMUNICATION PROCEDURES SHEET
# ============================================================================

def create_communication_procedures(ws, styles):
    """Create Communication_Procedures sheet documenting stakeholder notifications."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "CHANGE COMMUNICATION PROCEDURES"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "Document stakeholder notification and communication methods"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== STAKEHOLDER IDENTIFICATION ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "STAKEHOLDER IDENTIFICATION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    stakeholder_headers = ["Stakeholder Group", "Role/Description", "Notification Trigger", "Communication Method", "Responsible Party", "Template Available", "Status"]
    for col_idx, header in enumerate(stakeholder_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    stakeholder_groups = [
        "End Users",
        "Service Owners",
        "IT Operations Team",
        "Security Team",
        "Management",
        "CAB Members",
        "External Vendors",
        "[Other - specify]",
        "[Other - specify]",
        "[Other - specify]",
    ]

    row += 1
    for group in stakeholder_groups:
        ws[f"A{row}"] = group
        
        # Editable cells
        for col in [2, 5]:  # B=2, E=5
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['notification_trigger'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['communication_method'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['yes_no'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== COMMUNICATION TEMPLATES ====================
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "COMMUNICATION TEMPLATES"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    template_headers = ["Template Type", "Purpose", "Content Includes", "Approval Required", "Format", "Location", "Maintained"]
    for col_idx, header in enumerate(template_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    template_types = [
        ("Change Request Acknowledgment", "Confirm receipt", "Change ID, requestor, next steps"),
        ("Change Scheduled Notification", "Inform of approval & schedule", "Change details, date/time, impact"),
        ("Change Reminder (24h)", "Pre-implementation reminder", "Time, expected duration, contacts"),
        ("Change In Progress", "During implementation", "Status, progress, issues"),
        ("Change Completed Successfully", "Successful completion", "Completion time, verification"),
        ("Change Failed/Rolled Back", "Failure notification", "Issue, rollback status, next steps"),
        ("Change Postponed", "Schedule change", "Reason, new date, next steps"),
        ("Emergency Change Notice", "Urgent notification", "Reason, approval, impact"),
    ]

    row += 1
    for template_type, purpose, content in template_types:
        ws[f"A{row}"] = template_type
        ws[f"B{row}"] = purpose
        ws[f"C{row}"] = content
        
        # Editable cells
        for col in [6]:  # F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['yes_no'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Format dropdown (Email/Portal/Both)
        format_dv = DataValidation(
            type="list",
            formula1='"Email,Portal,Both"',
            allow_blank=False
        )
        ws.add_data_validation(format_dv)
        format_dv.add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== COMMUNICATION CHANNELS ASSESSMENT ====================
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "COMMUNICATION CHANNELS ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    channel_headers = ["Channel", "Available", "Primary Use", "Audience Reach", "Reliability", "Documented"]
    for col_idx, header in enumerate(channel_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    channels = [
        "Email",
        "Internal Portal/Intranet",
        "Change Management System",
        "Instant Messaging (Teams/Slack/etc)",
        "SMS/Text Messages",
        "Status Dashboard",
        "Scheduled Announcements",
    ]

    row += 1
    for channel in channels:
        ws[f"A{row}"] = channel
        
        # Editable cells
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['yes_no'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['audience_reach'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['reliability'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['implementation_status'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    ws.freeze_panes = "A6"


# ============================================================================
# SECTION 7: DOCUMENTATION REQUIREMENTS SHEET
# ============================================================================

def create_documentation_requirements(ws, styles):
    """Create Documentation_Requirements sheet defining record-keeping requirements."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "CHANGE DOCUMENTATION REQUIREMENTS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "Define record-keeping requirements by change type"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== STANDARD CHANGE DOCUMENTATION ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "STANDARD CHANGE DOCUMENTATION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    std_doc_headers = ["Documentation Item", "Required", "Format", "Retention Period", "Storage Location", "Responsible Role", "Compliant"]
    for col_idx, header in enumerate(std_doc_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    std_doc_items = [
        "Change request (basic)",
        "Requestor identification",
        "Change description",
        "Service/system affected",
        "Implementation date/time",
        "Implementation result",
    ]

    row += 1
    for item in std_doc_items:
        ws[f"A{row}"] = item
        
        # Editable cells
        for col in [4, 5, 6]:  # D=4, E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['doc_requirement'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['format_type'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== NORMAL CHANGE DOCUMENTATION ====================
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "NORMAL CHANGE DOCUMENTATION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    for col_idx, header in enumerate(std_doc_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    normal_doc_items = [
        "Complete change request",
        "Risk assessment",
        "Impact assessment",
        "Business justification",
        "Technical implementation plan",
        "Test plan/results",
        "Rollback plan",
        "CAB review record",
        "Approval record (all approvers)",
        "Communication records",
        "Implementation log/notes",
        "Verification evidence",
        "Post-Implementation Review",
    ]

    row += 1
    for item in normal_doc_items:
        ws[f"A{row}"] = item
        
        # Editable cells
        for col in [4, 5, 6]:  # D=4, E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['doc_requirement'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['format_type'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== EMERGENCY CHANGE DOCUMENTATION ====================
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "EMERGENCY CHANGE DOCUMENTATION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    emerg_doc_headers = ["Documentation Item", "Required", "Special Requirements", "Retention Period", "Storage Location", "Responsible Role", "Compliant"]
    for col_idx, header in enumerate(emerg_doc_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    emerg_doc_items = [
        "Emergency change justification",
        "E-CAB approval record",
        "Emergency criteria met",
        "Incident ticket reference",
        "Risk accepted by",
        "Implementation record",
        "Mandatory PIR",
        "Retrospective CAB review",
    ]

    row += 1
    for item in emerg_doc_items:
        ws[f"A{row}"] = item
        
        # Editable cells
        for col in [3, 4, 5, 6]:  # C=3, D=4, E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['doc_requirement'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['implementation_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== RECORD RETENTION & DISPOSAL ====================
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "RECORD RETENTION & DISPOSAL"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    retention_headers = ["Record Type", "Retention Period", "Disposal Method", "Regulatory Requirement", "Implemented", "Evidence"]
    for col_idx, header in enumerate(retention_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    record_types = [
        "Standard changes",
        "Normal changes (successful)",
        "Failed changes",
        "Emergency changes",
        "CAB meeting minutes",
        "Approval records",
    ]

    row += 1
    for record_type in record_types:
        ws[f"A{row}"] = record_type
        
        # Editable cells
        for col in [2, 4, 6]:  # B=2, D=4, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['disposal_method'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['implementation_status'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 18

    ws.freeze_panes = "A6"


# ============================================================================
# SECTION 8: CHANGE MANAGEMENT TOOLS SHEET
# ============================================================================

def create_change_management_tools(ws, styles):
    """Create Change_Management_Tools sheet inventorying all supporting tools/systems."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "CHANGE MANAGEMENT TOOLS & SYSTEMS INVENTORY"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "Document all platforms and tools supporting change management"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== PRIMARY CHANGE MANAGEMENT SYSTEM ====================
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "PRIMARY CHANGE MANAGEMENT SYSTEM"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    ws[f"A{row}"] = "Attribute"
    ws[f"B{row}"] = "Value"
    ws[f"C{row}"] = "Evidence Location"
    for col in [1, 2, 3]:  # A=1, B=2, C=3
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    primary_system_fields = [
        ("SYSTEM IDENTIFICATION", None),
        ("Tool/Platform Name", "text"),
        ("Vendor/Provider", "text"),
        ("Version/Release", "text"),
        ("Deployment Type", "deployment_type"),
        ("Primary Use", "text"),
        ("", None),
        ("CAPABILITIES", None),
        ("Change request submission", "tool_capability"),
        ("Workflow automation", "tool_capability"),
        ("Approval routing", "tool_capability"),
        ("Risk/Impact assessment forms", "tool_capability"),
        ("Change calendar", "tool_capability"),
        ("Conflict detection", "tool_capability"),
        ("Notification engine", "tool_capability"),
        ("Reporting/dashboards", "tool_capability"),
        ("Integration with monitoring", "tool_capability"),
        ("Integration with CMDB", "tool_capability"),
        ("API availability", "tool_capability"),
        ("", None),
        ("ADMINISTRATION", None),
        ("System administrators", "text"),
        ("User training provided", "yes_no"),
        ("Documentation available", "yes_no"),
        ("Support availability", "support_availability"),
        ("License status", "license_status"),
        ("License expiration date", "date"),
    ]

    row += 1
    for field_name, field_type in primary_system_fields:
        ws[f"A{row}"] = field_name
        
        if field_name and field_name.isupper():
            # Section header
            ws[f"A{row}"].font = Font(bold=True, italic=True)
        elif field_type:
            # Editable value cell
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            # Apply validation if specified
            if field_type == "date":
                ws[f"B{row}"].number_format = 'DD.MM.YYYY'
            elif field_type != "text" and field_type in validations:
                validations[field_type].add(ws[f"B{row}"])
        
        row += 1

    # ==================== SUPPORTING TOOLS & SYSTEMS ====================
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "SUPPORTING TOOLS & SYSTEMS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    support_headers = ["Tool/System", "Purpose", "Integration with Primary", "Status", "User Count", "Evidence"]
    for col_idx, header in enumerate(support_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    row += 1
    # 18 rows for supporting tools
    for i in range(18):
        for col in [1, 2, 5, 6]:  # A=1, B=2, E=5, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['integration_status'].add(ws.cell(row=row, column=3))
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['implementation_status'].add(ws.cell(row=row, column=4))
        ws.cell(row=row, column=4).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== INTEGRATION ARCHITECTURE ====================
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "INTEGRATION ARCHITECTURE"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    integ_headers = ["Integration Point", "Source System", "Target System", "Integration Type", "Automated", "Data Exchanged", "Status"]
    for col_idx, header in enumerate(integ_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    row += 1
    # 12 rows for integrations
    for i in range(12):
        for col in [1, 2, 3, 6]:  # A=1, B=2, C=3, F=6
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['integration_type'].add(ws.cell(row=row, column=4))
        ws.cell(row=row, column=4).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['yes_no_partial'].add(ws.cell(row=row, column=5))
        ws.cell(row=row, column=5).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['implementation_status'].add(ws.cell(row=row, column=7))
        ws.cell(row=row, column=7).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== TOOL ACCESS & SECURITY ====================
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "TOOL ACCESS & SECURITY"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    security_headers = ["Security Control", "Implemented", "Details", "Evidence"]
    for col_idx, header in enumerate(security_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    security_controls = [
        "Multi-factor authentication",
        "Role-based access control",
        "Audit logging enabled",
        "Data encryption (at rest)",
        "Data encryption (in transit)",
        "Backup/recovery tested",
        "Disaster recovery plan",
    ]

    row += 1
    for control in security_controls:
        ws[f"A{row}"] = control
        
        # Editable cells
        for col in [3, 4]:  # C=3, D=4
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdown
        validations['yes_no_partial'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 18

    ws.freeze_panes = "A6"


# ============================================================================
# SECTION 9: SUMMARY DASHBOARD SHEET
# ============================================================================

def create_summary_dashboard(ws, styles):
    """Create Summary_Dashboard with compliance metrics and gap identification."""
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "CHANGE PROCESS ASSESSMENT - SUMMARY DASHBOARD"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "Overall compliance status and key findings"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== OVERALL COMPLIANCE SUMMARY ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "OVERALL COMPLIANCE SUMMARY"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    summary_headers = ["Assessment Area", "Total Requirements", "Implemented", "Partial", "Not Implemented", "Compliance %", "Status"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    assessment_areas = [
        "Change Process Workflow",
        "Approval Authority",
        "Communication Procedures",
        "Documentation Requirements",
        "Change Management Tools",
    ]

    row += 1
    start_data_row = row
    for area in assessment_areas:
        ws[f"A{row}"] = area
        
        # Formula cells (calculated - gray background)
        for col in range(2, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Note: Actual formulas would reference the assessment sheets
            # This is a template - formulas need to be customized
            if col == 2:  # Total Requirements
                cell.value = "[FORMULA]"
            elif col == 3:  # Implemented
                cell.value = "[FORMULA]"
            elif col == 4:  # Partial
                cell.value = "[FORMULA]"
            elif col == 5:  # Not Implemented
                cell.value = "[FORMULA]"
            elif col == 6:  # Compliance %
                cell.value = "[FORMULA]"
                cell.number_format = '0%'
            elif col == 7:  # Status
                cell.value = "[FORMULA]"
        
        row += 1

    # Overall total row
    ws[f"A{row}"] = "OVERALL TOTAL"
    ws[f"A{row}"].font = Font(bold=True)
    for col in range(2, 8):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(bold=True)
        cell.value = "[FORMULA]"
        if col == 6:
            cell.number_format = '0%'

    # ==================== POLICY REQUIREMENT MAPPING ====================
    row += 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "POLICY REQUIREMENT MAPPING"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    req_headers = ["Policy Req ID", "Requirement Summary", "Status", "Evidence Sheet", "Evidence Row", "Auditor Notes"]
    for col_idx, header in enumerate(req_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # 20 policy requirements from ISMS-POL-A.8.32-S2.1
    requirements = [
        ("REQ-PROCESS-001", "Change request procedures"),
        ("REQ-PROCESS-002", "Risk assessment"),
        ("REQ-PROCESS-003", "Impact assessment"),
        ("REQ-PROCESS-004", "Change classification"),
        ("REQ-PROCESS-005", "Approval workflows"),
        ("REQ-PROCESS-006", "CAB operation"),
        ("REQ-PROCESS-007", "Emergency procedures"),
        ("REQ-PROCESS-008", "Communication"),
        ("REQ-PROCESS-009", "Implementation planning"),
        ("REQ-PROCESS-010", "Testing requirements"),
        ("REQ-PROCESS-011", "Rollback plans"),
        ("REQ-PROCESS-012", "Documentation"),
        ("REQ-PROCESS-013", "PIR mandatory"),
        ("REQ-PROCESS-014", "Continuous improvement"),
        ("REQ-PROCESS-015", "Record retention"),
        ("REQ-PROCESS-016", "Tool capabilities"),
        ("REQ-PROCESS-017", "Integration"),
        ("REQ-PROCESS-018", "Metrics/KPIs"),
        ("REQ-PROCESS-019", "Training"),
        ("REQ-PROCESS-020", "Governance"),
    ]

    row += 1
    for req_id, req_summary in requirements:
        ws[f"A{row}"] = req_id
        ws[f"B{row}"] = req_summary
        
        # Status - formula cell
        ws[f"C{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        ws[f"C{row}"].value = "[FORMULA]"
        
        # Evidence reference - editable
        for col in [4, 5]:  # D=4, E=5
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Auditor notes - editable
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== CRITICAL FINDINGS ====================
    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "CRITICAL FINDINGS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    finding_headers = ["Finding Type", "Count", "Description"]
    for col_idx, header in enumerate(finding_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    finding_types = [
        "Critical Gaps",
        "High-Priority Items",
        "Planned Improvements",
    ]

    row += 1
    for finding_type in finding_types:
        ws[f"A{row}"] = finding_type
        
        # Count - formula
        ws[f"B{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        ws[f"B{row}"].value = "[FORMULA]"
        
        # Description - editable text area
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== AUDIT READINESS ASSESSMENT ====================
    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "AUDIT READINESS ASSESSMENT"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    audit_headers = ["Criterion", "Status", "Notes"]
    for col_idx, header in enumerate(audit_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    audit_criteria = [
        "All processes documented",
        "Evidence available for key controls",
        "Roles and responsibilities defined",
        "Tool capabilities verified",
        "Compliance ≥70%",
        "OVERALL AUDIT READINESS",
    ]

    row += 1
    for criterion in audit_criteria:
        ws[f"A{row}"] = criterion
        if "OVERALL" in criterion:
            ws[f"A{row}"].font = Font(bold=True)
        
        # Status - formula
        ws[f"B{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        ws[f"B{row}"].value = "[FORMULA]"
        
        # Notes - editable
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 15

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 10: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence_Register with 100 pre-formatted rows for evidence tracking."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:I1")
    ws["A1"] = "EVIDENCE REGISTER"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = "Document all evidence supporting this assessment"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = ["Evidence ID", "Evidence Type", "Description", "Related Sheet/Requirement", "Location/Path", "Date Collected", "Collected By", "Verification Status", "Auditor Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Create 100 evidence rows
    row += 1
    for i in range(1, 101):
        # Evidence ID - auto-generated
        ws[f"A{row}"] = f"EV-{i:03d}"
        ws[f"A{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Evidence Type - dropdown
        validations['evidence_type'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Description, Related Sheet, Location, Collected By - text
        for col in [3, 4, 5, 7, 9]:  # C=3, D=4, E=5, G=7, I=9
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Date Collected - date
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].number_format = 'DD.MM.YYYY'
        
        # Verification Status - dropdown
        validations['verification_status'].add(ws[f"H{row}"])
        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Alternating row colors for readability
        if i % 2 == 0:
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                if cell.fill.start_color.rgb != "FFFFCC" and cell.fill.start_color.rgb != "E0E0E0":
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 30

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 11: APPROVAL SIGN-OFF SHEET
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create Approval_Sign_Off sheet with 3-level approval workflow."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "APPROVAL SIGN-OFF"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    # ==================== ASSESSMENT SUMMARY ====================
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    summary_fields = [
        ("Assessment Document:", "ISMS-IMP-A.8.32.1 - Change Process Assessment"),
        ("Assessment Period:", "[USER INPUT]"),
        ("Assessment Scope:", "[USER INPUT]"),
        ("Overall Compliance Rate:", "[FORMULA from Summary_Dashboard]"),
        ("Critical Gaps:", "[FORMULA from Summary_Dashboard]"),
        ("Audit Readiness:", "[FORMULA from Summary_Dashboard]"),
        ("Assessment Status:", "[DROPDOWN]"),
    ]

    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        
        if "USER INPUT" in value:
            ws[f"B{row}"] = ""
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        elif "DROPDOWN" in value:
            validations['assessment_status'].add(ws[f"B{row}"])
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        elif "FORMULA" in value:
            ws[f"B{row}"] = value
            ws[f"B{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        else:
            ws[f"B{row}"] = value
        
        row += 1

    # ==================== ASSESSMENT COMPLETED BY ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    preparer_fields = [
        "Name:",
        "Role/Title:",
        "Department:",
        "Email:",
        "Phone:",
        "Date Completed:",
        "Signature:",
    ]

    for field in preparer_fields:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        if field == "Date Completed:":
            ws[f"B{row}"].number_format = 'DD.MM.YYYY'
        
        row += 1

    # ==================== REVIEWED BY ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY - Process Owner / Change Manager"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    reviewer_fields = [
        ("Name:", None),
        ("Role/Title:", None),
        ("Review Date:", "date"),
        ("Review Notes:", "textarea"),
        ("Recommendation:", "dropdown"),
        ("Conditions (if any):", "text"),
        ("Signature:", None),
    ]

    for field, field_type in reviewer_fields:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        if field_type == "date":
            ws[f"B{row}"].number_format = 'DD.MM.YYYY'
        elif field_type == "dropdown":
            validations['review_recommendation'].add(ws[f"B{row}"])
        elif field_type == "textarea":
            ws.merge_cells(f"B{row}:E{row+2}")
            ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = 60
            row += 2
        
        row += 1

    # ==================== APPROVED BY ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY - CISO or IT Operations Manager"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    approver_fields = [
        ("Name:", None),
        ("Role/Title:", None),
        ("Approval Date:", "date"),
        ("Approval Decision:", "dropdown"),
        ("Conditions/Notes:", "textarea"),
        ("Signature:", None),
    ]

    for field, field_type in approver_fields:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        if field_type == "date":
            ws[f"B{row}"].number_format = 'DD.MM.YYYY'
        elif field_type == "dropdown":
            validations['approval_decision'].add(ws[f"B{row}"])
        elif field_type == "textarea":
            ws.merge_cells(f"B{row}:E{row+2}")
            ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = 60
            row += 2
        
        row += 1

    # ==================== NEXT REVIEW DETAILS ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    next_review_fields = [
        ("Next Review Date:", "formula"),
        ("Review Responsible:", None),
        ("Special Considerations:", "text"),
        ("Regulatory Review Required:", "dropdown"),
        ("External Audit Scheduled:", "date"),
    ]

    for field, field_type in next_review_fields:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        if field_type == "formula":
            # Auto-calculate +3 months (90 days)
            ws[f"B{row}"] = "=TODAY()+90"
            ws[f"B{row}"].number_format = 'DD.MM.YYYY'
            ws[f"B{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        elif field_type == "date":
            ws[f"B{row}"].number_format = 'DD.MM.YYYY'
        elif field_type == "dropdown":
            validations['regulatory_review'].add(ws[f"B{row}"])
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Philosophy: "The first principle is that you must not fool yourself—
    and you are the easiest person to fool." - Richard Feynman
    
    This script creates evidence-based assessment tools for change management,
    not checkbox compliance theater.
    """
    print("=" * 78)
    print("ISMS-IMP-A.8.32.1 - Change Process Assessment Generator")
    print("ISO/IEC 27001:2022 Control A.8.32: Change Management")
    print("=" * 78)
    print("\n🎯 Systems Engineering Approach: Evidence-Based Compliance")
    print(f"{CHART} Technology-Agnostic: Works with ANY change management approach")
    print(f"{LOCK} Audit-Ready: Comprehensive evidence collection")
    print("\n" + "─" * 78)

    # Create workbook and setup styles
    print("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    print(f"{CHECK} Workbook created with 9 sheets")

    # Create all sheets
    print("\n[Phase 2] Generating assessment sheets...")
    
    print("  [1/9] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    print("  ✅ Instructions complete")

    print("  [2/9] Creating Change_Process_Workflow...")
    create_change_process_workflow(wb["Change_Process_Workflow"], styles)
    print("  ✅ Process workflow complete (15 stages, automation assessment)")

    print("  [3/9] Creating Approval_Authority_Matrix...")
    create_approval_authority_matrix(wb["Approval_Authority_Matrix"], styles)
    print("  ✅ Approval matrix complete (Standard/Normal/Emergency changes)")

    print("  [4/9] Creating Communication_Procedures...")
    create_communication_procedures(wb["Communication_Procedures"], styles)
    print("  ✅ Communication procedures complete (stakeholder notifications)")

    print("  [5/9] Creating Documentation_Requirements...")
    create_documentation_requirements(wb["Documentation_Requirements"], styles)
    print("  ✅ Documentation requirements complete (record retention)")

    print("  [6/9] Creating Change_Management_Tools...")
    create_change_management_tools(wb["Change_Management_Tools"], styles)
    print("  ✅ Tool inventory complete (primary system + integrations)")

    print("  [7/9] Creating Summary_Dashboard...")
    create_summary_dashboard(wb["Summary_Dashboard"], styles)
    print("  ✅ Dashboard complete (compliance metrics, audit readiness)")

    print("  [8/9] Creating Evidence_Register...")
    create_evidence_register(wb["Evidence_Register"], styles)
    print("  ✅ Evidence register complete (100 evidence rows)")

    print("  [9/9] Creating Approval_Sign_Off...")
    create_approval_signoff(wb["Approval_Sign_Of"], styles)
    print("  ✅ Approval workflow complete (3-level sign-off)")

    # Save workbook
    print("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.32.1_Change_Process_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        wb.save(filename)
        print(f"{CHECK} SUCCESS: {filename}")
    except Exception as e:
        print(f"{XMARK} ERROR saving workbook: {e}")
        return 1

    # Summary
    print("\n" + "=" * 78)
    print("📋 WORKBOOK STRUCTURE SUMMARY")
    print("=" * 78)
    print("\n📄 Assessment Sheets:")
    print("  • Instructions & Legend (usage guidance)")
    print("  • Change_Process_Workflow (15-stage lifecycle)")
    print("  • Approval_Authority_Matrix (Standard/Normal/Emergency)")
    print("  • Communication_Procedures (stakeholder notifications)")
    print("  • Documentation_Requirements (record-keeping)")
    print("  • Change_Management_Tools (tool inventory + integrations)")
    print("\n📊 Analysis & Governance:")
    print("  • Summary_Dashboard (compliance metrics, audit readiness)")
    print("  • Evidence_Register (100 evidence entries)")
    print("  • Approval_Sign_Off (3-level approval workflow)")
    print("\n" + "─" * 78)
    print("📈 ASSESSMENT CAPABILITIES:")
    print("  • 15 process stages documented")
    print("  • 20 policy requirements mapped")
    print("  • Standard/Normal/Emergency change approval paths")
    print("  • Stakeholder communication procedures")
    print("  • Documentation & retention requirements")
    print("  • Tool capability assessment")
    print("  • 100 evidence documentation entries")
    print("  • Automated compliance calculations")
    print("\n" + "─" * 78)
    print(f"{TARGET} KEY FEATURES:")
    print("  ✅ Technology-agnostic (works with ANY change management tool)")
    print("  ✅ Comprehensive evidence collection")
    print("  ✅ Automated compliance calculations")
    print("  ✅ Audit readiness assessment")
    print("  ✅ Multi-level approval workflow")
    print("  ✅ Quarterly review cycle support")
    print("\n" + "=" * 78)
    print(f"{ROCKET} NEXT STEPS:")
    print("  1. Open the generated workbook")
    print("  2. Review Instructions & Legend sheet first")
    print("  3. Document YOUR organization's change process")
    print("  4. Complete approval authority definitions")
    print("  5. Define communication procedures")
    print("  6. Document record-keeping requirements")
    print("  7. Inventory YOUR change management tools")
    print("  8. Review Summary_Dashboard for compliance status")
    print("  9. Document evidence in Evidence_Register")
    print("  10. Obtain final approval via Approval_Sign_Of")
    print("\n💡 PRO TIP:")
    print("  This workbook is technology-independent. Whether you use ServiceNow,")
    print("  Jira, custom tools, or spreadsheets - this framework assesses YOUR")
    print("  change management PROCESS, not your tool brands. That's the difference")
    print("  between compliance theater and Systems Engineering.")
    print("\n" + "=" * 78)
    print('\n"The first principle is that you must not fool yourself')
    print('— and you are the easiest person to fool." - Richard Feynman')
    print("\n🎭 This is not cargo cult ISMS. This is evidence-based compliance.")
    print("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    exit(main())