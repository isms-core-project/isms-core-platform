#!/usr/bin/env python3
"""
Excel Sanity Check for ISMS-IMP-A.8.10 (Complete Framework)
Validates all 5 deletion assessment workbooks (A.8.10.1-5)

This script checks for:
- Workbook structure (correct sheets present)
- Cell reference verification (external links point to correct cells)
- Data validations (properly configured dropdowns)
- Formula syntax (no #REF!, #NAME? errors)
- A.8.10.4 special case (Verification Dashboard vs Summary Dashboard)
- A.8.10.5 external link formula validation
- Column widths appropriate
- Freeze panes set correctly

CRITICAL: This script validates the CORRECTED cell references:
    A.8.10.1-3: 'Summary Dashboard'!G10
    A.8.10.4:   'Verification Dashboard'!B9 (DIFFERENT!)

Usage:
    python3 excel_sanity_check_a810_complete.py [workbook_directory]
    
    If no directory specified, checks current directory.

"""

import sys
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import re


# ============================================================================
# CONFIGURATION
# ============================================================================

# Expected workbook structures (based on actual generator scripts)
WORKBOOK_SPECS = {
    "A.8.10.1": {
        "pattern": "ISMS-IMP-A.8.10.1",
        "name": "Retention & Deletion Triggers Assessment",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Data Category Registry",
            "3. Retention Schedule Compliance",
            "4. Deletion Trigger Configuration",
            "5. Legal Hold Management",
            "6. Data Subject Requests",
            "Summary Dashboard",
            "Evidence Register",
            "Approval Sign-Off"
        ],
        "min_sheets": 9,
        "dashboard_sheet": "Summary Dashboard",
        "key_cells": {
            "compliance_pct": "G10",
            "non_compliant": "E10",
            "total_items": "B10"
        }
    },
    "A.8.10.2": {
        "pattern": "ISMS-IMP-A.8.10.2",
        "name": "Deletion Methods Assessment",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Physical Media Deletion",
            "3. Cloud Storage Deletion",
            "4. Database & Application Deletion",
            "5. Mobile & Endpoint Deletion",
            "6. Deletion Tool Validation",
            "Summary Dashboard",
            "Evidence Register",
            "Approval Sign-Off"
        ],
        "min_sheets": 9,
        "dashboard_sheet": "Summary Dashboard",
        "key_cells": {
            "compliance_pct": "G10",
            "non_compliant": "E10",
            "total_items": "B10"
        }
    },
    "A.8.10.3": {
        "pattern": "ISMS-IMP-A.8.10.3",
        "name": "Third-Party & Cloud Deletion Assessment",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Cloud Provider Assessment",
            "3. SaaS Application Deletion",
            "4. Vendor Contract Review",
            "5. Subprocessor Mapping",
            "6. Vendor Performance Tracking",
            "Summary Dashboard",
            "Evidence Register",
            "Approval Sign-Off"
        ],
        "min_sheets": 9,
        "dashboard_sheet": "Summary Dashboard",
        "key_cells": {
            "compliance_pct": "G10",
            "non_compliant": "E10",
            "total_items": "B10"
        }
    },
    "A.8.10.4": {
        "pattern": "ISMS-IMP-A.8.10.4",
        "name": "Verification & Evidence Assessment",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Deletion Logging",
            "3. Verification Testing",
            "4. Evidence Repository",
            "5. Certificate Management",
            "6. Audit Trail",
            "Verification Dashboard",  # DIFFERENT NAME!
            "Evidence Register",
            "Approval Sign-Off"
        ],
        "min_sheets": 9,
        "dashboard_sheet": "Verification Dashboard",  # CRITICAL: Not "Summary Dashboard"!
        "key_cells": {
            "compliance_pct": "B9",  # DIFFERENT CELL!
            "critical_gaps": "B8",
            "effectiveness": "B6"
        }
    },
    "A.8.10.5": {
        "pattern": "ISMS-IMP-A.8.10.5",
        "name": "Compliance Dashboard (External Links)",
        "expected_sheets": [
            "Instructions",
            "Overall A.8.10 Compliance",
            "Retention Schedule Health",
            "Deletion Method Effectiveness",
            "Third-Party Deletion Performance",
            "Verification & Evidence Quality",
            "Critical Gaps Dashboard",
            "Trend Analysis",
            "Executive Summary"
        ],
        "min_sheets": 9,
        "has_external_links": True,
        "expected_formulas": [
            "='[ISMS-IMP-A.8.10.1.xlsx]Summary Dashboard'!G10",
            "='[ISMS-IMP-A.8.10.2.xlsx]Summary Dashboard'!G10",
            "='[ISMS-IMP-A.8.10.3.xlsx]Summary Dashboard'!G10",
            "='[ISMS-IMP-A.8.10.4.xlsx]Verification Dashboard'!B9"  # DIFFERENT!
        ]
    }
}

# Color codes for terminal output
GREEN = '\033[92m'
YELLOW = '\033[93m'
RED = '\033[91m'
BLUE = '\033[94m'
BOLD = '\033[1m'
RESET = '\033[0m'


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def print_header(text):
    """Print section header."""
    print(f"\n{BLUE}{BOLD}{'=' * 80}{RESET}")
    print(f"{BLUE}{BOLD}{text}{RESET}")
    print(f"{BLUE}{BOLD}{'=' * 80}{RESET}\n")


def print_success(text):
    """Print success message."""
    print(f"{GREEN}✅ {text}{RESET}")


def print_warning(text):
    """Print warning message."""
    print(f"{YELLOW}⚠️  {text}{RESET}")


def print_error(text):
    """Print error message."""
    print(f"{RED}❌ {text}{RESET}")


def print_info(text):
    """Print info message."""
    print(f"   {text}")


def find_workbook(directory, pattern):
    """Find workbook matching pattern in directory."""
    path = Path(directory)
    matches = list(path.glob(f"{pattern}*.xlsx"))
    
    # Filter out temp files
    matches = [m for m in matches if not m.name.startswith("~$")]
    
    if not matches:
        return None
    
    if len(matches) > 1:
        print_warning(f"Multiple workbooks found for {pattern}, using most recent")
        matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    
    return matches[0]


def check_workbook_structure(wb, assessment_key, spec):
    """Check if workbook has correct sheet structure."""
    issues = []
    warnings = []
    
    actual_sheets = wb.sheetnames
    expected_sheets = spec["expected_sheets"]
    
    # Check sheet count
    if len(actual_sheets) < spec["min_sheets"]:
        issues.append(f"Only {len(actual_sheets)} sheets found, expected at least {spec['min_sheets']}")
    else:
        print_success(f"Sheet count: {len(actual_sheets)} sheets")
    
    # Check for expected sheets
    missing_sheets = []
    for sheet_name in expected_sheets:
        if sheet_name not in actual_sheets:
            missing_sheets.append(sheet_name)
    
    if missing_sheets:
        issues.append(f"Missing sheets: {', '.join(missing_sheets)}")
    else:
        print_success(f"All {len(expected_sheets)} expected sheets present")
    
    # Check for unexpected sheets
    unexpected_sheets = []
    for sheet_name in actual_sheets:
        if sheet_name not in expected_sheets:
            unexpected_sheets.append(sheet_name)
    
    if unexpected_sheets:
        warnings.append(f"Unexpected sheets: {', '.join(unexpected_sheets)}")
    
    return issues, warnings


def check_dashboard_cells(wb, assessment_key, spec):
    """Check if key dashboard cells exist and are in correct locations."""
    issues = []
    warnings = []
    
    if "dashboard_sheet" not in spec:
        return issues, warnings  # A.8.10.5 doesn't have this check
    
    dashboard_sheet = spec["dashboard_sheet"]
    
    if dashboard_sheet not in wb.sheetnames:
        issues.append(f"Dashboard sheet '{dashboard_sheet}' not found")
        return issues, warnings
    
    ws = wb[dashboard_sheet]
    key_cells = spec.get("key_cells", {})
    
    print_info(f"\nVerifying key cells in '{dashboard_sheet}':")
    
    for cell_name, cell_ref in key_cells.items():
        try:
            cell = ws[cell_ref]
            
            # Check if cell exists and has something (value or formula)
            if cell.value is not None or cell.data_type == 'f':
                print_success(f"  {cell_name}: {cell_ref} exists")
            else:
                warnings.append(f"{dashboard_sheet}: Cell {cell_ref} ({cell_name}) is empty")
        except Exception as e:
            issues.append(f"{dashboard_sheet}: Cannot access cell {cell_ref} ({cell_name}): {str(e)}")
    
    # Special check for A.8.10.4: Verify it's NOT using "Summary Dashboard"
    if assessment_key == "A.8.10.4":
        if "Summary Dashboard" in wb.sheetnames:
            issues.append("A.8.10.4 has 'Summary Dashboard' sheet (should be 'Verification Dashboard')")
        else:
            print_success("A.8.10.4 correctly uses 'Verification Dashboard' (not 'Summary Dashboard')")
    
    return issues, warnings


def check_external_link_formulas(wb, assessment_key, spec):
    """Check A.8.10.5 dashboard for correct external link formulas."""
    issues = []
    warnings = []
    
    if assessment_key != "A.8.10.5":
        return issues, warnings  # Only check A.8.10.5
    
    if "Overall A.8.10 Compliance" not in wb.sheetnames:
        issues.append("A.8.10.5: Missing 'Overall A.8.10 Compliance' sheet")
        return issues, warnings
    
    ws = wb["Overall A.8.10 Compliance"]
    
    print_info("\nVerifying external link formulas in A.8.10.5:")
    
    # Check cells B8:B11 for external link formulas
    expected_patterns = {
        "B8": r".*ISMS-IMP-A\.8\.10\.1\.xlsx.*Summary Dashboard.*G10",
        "B9": r".*ISMS-IMP-A\.8\.10\.2\.xlsx.*Summary Dashboard.*G10",
        "B10": r".*ISMS-IMP-A\.8\.10\.3\.xlsx.*Summary Dashboard.*G10",
        "B11": r".*ISMS-IMP-A\.8\.10\.4\.xlsx.*Verification Dashboard.*B9"  # CRITICAL CHECK!
    }
    
    for cell_ref, pattern in expected_patterns.items():
        cell = ws[cell_ref]
        
        if cell.data_type == 'f':  # Formula
            formula = str(cell.value)
            
            if re.search(pattern, formula, re.IGNORECASE):
                print_success(f"  {cell_ref}: External link formula verified")
            else:
                issues.append(f"A.8.10.5 {cell_ref}: Formula doesn't match expected pattern")
                print_error(f"    Expected pattern: {pattern}")
                print_error(f"    Actual formula: {formula}")
        else:
            warnings.append(f"A.8.10.5 {cell_ref}: Not a formula (expected external link)")
    
    # Special emphasis on B11 (A.8.10.4 reference)
    cell_b11 = ws["B11"]
    if cell_b11.data_type == 'f':
        formula = str(cell_b11.value)
        
        if "Verification Dashboard" in formula and "B9" in formula:
            print_success("  ✨ A.8.10.4 reference CORRECTLY uses 'Verification Dashboard'!B9")
        elif "Summary Dashboard" in formula:
            issues.append("CRITICAL: A.8.10.5 B11 references 'Summary Dashboard' (should be 'Verification Dashboard')")
        else:
            warnings.append("A.8.10.5 B11: Cannot determine dashboard sheet name in formula")
    
    return issues, warnings


def check_data_validations(ws, sheet_name):
    """Check if data validations are properly configured."""
    issues = []
    warnings = []
    
    # Check if any validations exist
    if not ws.data_validations.dataValidation:
        if sheet_name not in ["Instructions & Legend", "Instructions", "Summary Dashboard", 
                              "Verification Dashboard", "Evidence Register"]:
            warnings.append(f"{sheet_name}: No data validations found")
        return issues, warnings
    
    validation_count = len(ws.data_validations.dataValidation)
    print_info(f"{sheet_name}: {validation_count} data validations configured")
    
    # Check each validation has cells assigned
    for idx, validation in enumerate(ws.data_validations.dataValidation, 1):
        if not validation.cells:
            issues.append(f"{sheet_name}: Validation #{idx} has no cells assigned")
    
    return issues, warnings


def check_formulas(ws, sheet_name):
    """Check for formula errors."""
    issues = []
    warnings = []
    
    formula_count = 0
    error_count = 0
    external_link_count = 0
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Formula
                formula_count += 1
                formula_str = str(cell.value)
                
                # Check for external links
                if '.xlsx]' in formula_str or '.xlsm]' in formula_str:
                    external_link_count += 1
                
                # Check for common error indicators
                if any(err in formula_str.upper() for err in ['#REF!', '#NAME?', '#VALUE!', '#DIV/0!']):
                    error_count += 1
                    issues.append(f"{sheet_name}: Formula error in {cell.coordinate}: {formula_str[:50]}")
    
    if formula_count > 0:
        print_info(f"{sheet_name}: {formula_count} formulas found ({external_link_count} external links)")
        if error_count == 0:
            print_success(f"{sheet_name}: No formula errors detected")
        else:
            print_error(f"{sheet_name}: {error_count} formula errors found")
    
    return issues, warnings


def check_freeze_panes(ws, sheet_name):
    """Check if freeze panes are set."""
    issues = []
    warnings = []
    
    if ws.freeze_panes:
        print_info(f"{sheet_name}: Freeze panes set at {ws.freeze_panes}")
    else:
        if sheet_name not in ["Summary Dashboard", "Verification Dashboard", "Instructions", "Instructions & Legend"]:
            warnings.append(f"{sheet_name}: No freeze panes set")
    
    return issues, warnings


def check_column_widths(ws, sheet_name):
    """Check if column widths are reasonable."""
    issues = []
    warnings = []
    
    narrow_columns = []
    wide_columns = []
    
    for col_letter, dimension in ws.column_dimensions.items():
        if dimension.width:
            if dimension.width < 8:
                narrow_columns.append(f"{col_letter}:{dimension.width}")
            elif dimension.width > 55:
                wide_columns.append(f"{col_letter}:{dimension.width}")
    
    if narrow_columns:
        warnings.append(f"{sheet_name}: Very narrow columns: {', '.join(narrow_columns)}")
    
    if wide_columns:
        warnings.append(f"{sheet_name}: Very wide columns: {', '.join(wide_columns)}")
    
    return issues, warnings


def validate_workbook(filepath, assessment_key, spec):
    """Validate a single workbook."""
    
    print_header(f"Checking {spec['name']} ({assessment_key})")
    print_info(f"File: {filepath.name}")
    
    try:
        # Load workbook
        wb = load_workbook(filepath, data_only=False)
        
        all_issues = []
        all_warnings = []
        
        # Check structure
        print_info("\nValidating workbook structure...")
        issues, warnings = check_workbook_structure(wb, assessment_key, spec)
        all_issues.extend(issues)
        all_warnings.extend(warnings)
        
        # Check dashboard cells (for A.8.10.1-4)
        if "dashboard_sheet" in spec:
            issues, warnings = check_dashboard_cells(wb, assessment_key, spec)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
        
        # Check external links (for A.8.10.5)
        if spec.get("has_external_links"):
            issues, warnings = check_external_link_formulas(wb, assessment_key, spec)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
        
        # Check each sheet
        print_info("\nValidating individual sheets...")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Data validations
            issues, warnings = check_data_validations(ws, sheet_name)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
            
            # Formulas
            issues, warnings = check_formulas(ws, sheet_name)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
            
            # Freeze panes
            issues, warnings = check_freeze_panes(ws, sheet_name)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
            
            # Column widths
            issues, warnings = check_column_widths(ws, sheet_name)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
        
        # Summary
        print_info("\n" + "=" * 60)
        if len(all_issues) == 0 and len(all_warnings) == 0:
            print_success(f"{assessment_key}: VALIDATION PASSED - No issues found")
            return True, 0, 0
        else:
            if len(all_issues) > 0:
                print_error(f"{assessment_key}: {len(all_issues)} critical issues found")
                for issue in all_issues:
                    print_error(f"  • {issue}")
            
            if len(all_warnings) > 0:
                print_warning(f"{assessment_key}: {len(all_warnings)} warnings found")
                for warning in all_warnings:
                    print_warning(f"  • {warning}")
            
            return len(all_issues) == 0, len(all_issues), len(all_warnings)
    
    except Exception as e:
        print_error(f"Failed to validate workbook: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, 1, 0


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main():
    """Main validation function."""
    
    # Determine directory
    if len(sys.argv) > 1:
        directory = sys.argv[1]
    else:
        directory = "."
    
    if not os.path.isdir(directory):
        print_error(f"Directory not found: {directory}")
        sys.exit(1)
    
    print_header("ISMS-IMP-A.8.10 Complete Framework Validation")
    print_info(f"Checking directory: {os.path.abspath(directory)}")
    print_info(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    print_info("\nCRITICAL: Verifying CORRECTED cell references:")
    print_info("  A.8.10.1-3: 'Summary Dashboard'!G10")
    print_info("  A.8.10.4:   'Verification Dashboard'!B9 (DIFFERENT!)")
    
    # Find and validate each assessment
    results = {}
    total_issues = 0
    total_warnings = 0
    
    for assessment_key, spec in WORKBOOK_SPECS.items():
        filepath = find_workbook(directory, spec["pattern"])
        
        if filepath:
            passed, issues, warnings = validate_workbook(filepath, assessment_key, spec)
            results[assessment_key] = {
                "found": True,
                "passed": passed,
                "issues": issues,
                "warnings": warnings
            }
            total_issues += issues
            total_warnings += warnings
        else:
            print_header(f"Checking {spec['name']} ({assessment_key})")
            print_warning(f"Workbook not found: {spec['pattern']}*.xlsx")
            results[assessment_key] = {
                "found": False,
                "passed": None,
                "issues": 0,
                "warnings": 0
            }
    
    # Final summary
    print_header("VALIDATION SUMMARY")
    
    found_count = sum(1 for r in results.values() if r["found"])
    passed_count = sum(1 for r in results.values() if r["passed"] == True)
    
    print_info(f"Workbooks found: {found_count} / 5")
    print_info(f"Workbooks passed: {passed_count} / {found_count}")
    print_info(f"Total critical issues: {total_issues}")
    print_info(f"Total warnings: {total_warnings}")
    
    print("\nAssessment Status:")
    for assessment_key, result in results.items():
        spec = WORKBOOK_SPECS[assessment_key]
        if result["found"]:
            if result["passed"]:
                print_success(f"{assessment_key} ({spec['name']}): PASSED")
            else:
                print_error(f"{assessment_key} ({spec['name']}): FAILED ({result['issues']} issues, {result['warnings']} warnings)")
        else:
            print_warning(f"{assessment_key} ({spec['name']}): NOT FOUND")
    
    # Special note about A.8.10.4
    if results.get("A.8.10.4", {}).get("found") and results["A.8.10.4"]["passed"]:
        print_info("\n✨ A.8.10.4 'Verification Dashboard' verified!")
    
    # Special note about A.8.10.5
    if results.get("A.8.10.5", {}).get("found") and results["A.8.10.5"]["passed"]:
        print_info("✨ A.8.10.5 external links verified!")
    
    # Exit code
    if total_issues > 0:
        print(f"\n{RED}{BOLD}VALIDATION FAILED{RESET}")
        print_info("Please fix critical issues before using workbooks in production")
        print_info("\nCommon fixes:")
        print_info("  - Regenerate workbooks with corrected generator scripts")
        print_info("  - Verify A.8.10.4 uses 'Verification Dashboard' sheet name")
        print_info("  - Verify A.8.10.5 references 'Verification Dashboard'!B9 for A.8.10.4")
        sys.exit(1)
    elif found_count == 0:
        print(f"\n{YELLOW}{BOLD}NO WORKBOOKS FOUND{RESET}")
        print_info("Generate workbooks first using the Python generator scripts:")
        print_info("  python3 generate_a810_1_retention_triggers.py")
        print_info("  python3 generate_a810_2_deletion_methods.py")
        print_info("  python3 generate_a810_3_third_party_cloud.py")
        print_info("  python3 generate_a810_4_verification_evidence.py")
        print_info("  python3 generate_a810_5_compliance_dashboard.py")
        sys.exit(2)
    else:
        print(f"\n{GREEN}{BOLD}VALIDATION PASSED{RESET}")
        if total_warnings > 0:
            print_info(f"Note: {total_warnings} warnings found (non-critical)")
        print_info("All found workbooks are ready for use")
        print_info("\nNext steps:")
        print_info("  1. Run: python3 normalize_assessment_files_a810.py")
        print_info("  2. Place dashboard in Dashboard_Sources folder")
        print_info("  3. Open dashboard and click 'Update Links'")
        sys.exit(0)


if __name__ == "__main__":
    main()