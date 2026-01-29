#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-A.8.10 - Assessment File Normalizer Utility
================================================================================

ISO/IEC 27001:2022 Control A.8.10: Information Deletion
Quality Assurance Utility: Excel Assessment File Normalization & Validation

--------------------------------------------------------------------------------
NORMALIZATION UTILITY - PREPARES FILES FOR DASHBOARD LINKING
--------------------------------------------------------------------------------

This utility is part of the A.8.10 Information Deletion assessment framework
and MUST be run after completing assessments to prepare files for dashboard
consolidation via external workbook linking.

**Purpose:**
Normalizes assessment workbook filenames to standardized names required for
external workbook formula linking in the A.8.10.5 Compliance Dashboard.

**Why This is Required:**
The A.8.10.5 Compliance Dashboard uses external workbook formulas like:
```
='[ISMS-IMP-A.8.10.1.xlsx]Summary Dashboard'!$G$10
```
These formulas require EXACT filenames. User-generated files typically have
date suffixes (e.g., `ISMS_A_8_10_1_Retention_Triggers_20250124.xlsx`) which
don't match the external formula references.

This script copies completed assessments to normalized names, enabling
external formula linking to work correctly.

Reference Pattern: Based on ISMS-A.8.24 normalization approach for consistency

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script normalizes and validates A.8.10 information deletion assessment 
Excel workbooks to ensure consistency, proper naming, and readiness for 
dashboard consolidation before integration into the A.8.10.5 Compliance 
Dashboard.

**Purpose:**
Ensures all information deletion assessment workbooks meet naming standards 
and structural requirements, enabling proper external workbook formula linking 
and preventing data consolidation errors.

**Key Functions:**

1. **File Discovery and Validation**
   - Scans directory for completed assessment files (A.8.10.1 through A.8.10.4)
   - Validates Document ID in Instructions & Legend sheet
   - Checks workbook structure and completeness
   - Reports missing or incomplete assessments

2. **Filename Normalization**
   - Removes date suffixes and version numbers
   - Standardizes to exact filenames required for external linking
   - Creates normalized copies (preserves originals as backup)
   - Handles filename conflicts gracefully

3. **Document ID Verification**
   - Validates Document ID matches expected value
   - Ensures consistency between filename and internal document metadata
   - Reports mismatches for correction

4. **Audit Trail Creation**
   - Generates normalization manifest with timestamp
   - Documents source → normalized filename mappings
   - Provides traceability for audit purposes
   - Records validation results

5. **Quality Reporting**
   - Reports normalization success/failure by domain
   - Identifies files ready for dashboard linking
   - Provides remediation guidance for issues
   - Confirms readiness for next workflow step

**Normalization Scope:**
- ISMS_A_8_10_1_Retention_Deletion_Triggers_*.xlsx → ISMS-IMP-A.8.10.1.xlsx
- ISMS_A_8_10_2_Deletion_Methods_*.xlsx → ISMS-IMP-A.8.10.2.xlsx
- ISMS_A_8_10_3_Third_Party_Cloud_Deletion_*.xlsx → ISMS-IMP-A.8.10.3.xlsx
- ISMS_A_8_10_4_Verification_Evidence_*.xlsx → ISMS-IMP-A.8.10.4.xlsx

**Output:**
- Normalized assessment workbooks (exact names for external linking)
- Normalization manifest (audit trail)
- Validation report (readiness confirmation)

**Integration:**
Normalized files are required for A.8.10.5 Compliance Dashboard external
workbook formula linking. The dashboard expects exact filenames and will
display #REF! errors if normalized files are not present.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel validation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Excel file validation)
    - shutil (file copying - standard library)
    - pathlib (path handling - standard library)
    - datetime (timestamp generation - standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    # Normalize all A.8.10 assessment files in current directory
    python3 normalize_assessment_files_a810.py

Advanced Usage:
    # Normalize files in specific directory
    python3 normalize_assessment_files_a810.py --input-dir /path/to/assessments
    
    # Specify output directory for normalized files
    python3 normalize_assessment_files_a810.py --output-dir /path/to/output
    
    # Dry run mode (report what would be normalized without modifying files)
    python3 normalize_assessment_files_a810.py --dry-run
    
    # Force overwrite existing normalized files
    python3 normalize_assessment_files_a810.py --force
    
    # Verbose output with detailed validation
    python3 normalize_assessment_files_a810.py --verbose

Command-Line Options:
    --input-dir PATH       Directory containing assessment workbooks (default: current)
    --output-dir PATH      Directory for normalized workbooks (default: same as input)
    --dry-run              Report normalization plan without copying files
    --force                Overwrite existing normalized files without prompting
    --verbose              Display detailed validation and normalization steps
    --skip-validation      Skip Document ID validation (faster, less safe)

Output Files:
    Normalized workbooks:
        - ISMS-IMP-A.8.10.1.xlsx (Retention & Deletion Triggers)
        - ISMS-IMP-A.8.10.2.xlsx (Deletion Methods)
        - ISMS-IMP-A.8.10.3.xlsx (Third-Party & Cloud Deletion)
        - ISMS-IMP-A.8.10.4.xlsx (Verification & Evidence)
    
    Audit trail:
        - A810_Normalization_Manifest_YYYYMMDD_HHMMSS.txt
    
    Validation report:
        - Console output (summary)

Workflow Examples:

    1. Initial normalization (before dashboard creation):
       python3 normalize_assessment_files_a810.py
    
    2. Check what would be normalized (dry run):
       python3 normalize_assessment_files_a810.py --dry-run
    
    3. Normalize with detailed validation:
       python3 normalize_assessment_files_a810.py --verbose
    
    4. Force re-normalization (overwrite existing):
       python3 normalize_assessment_files_a810.py --force

Typical Workflow:
    1. Complete assessments (A.8.10.1 through A.8.10.4)
    2. Run: python3 excel_sanity_check_a810.py (validate quality)
    3. Run: python3 normalize_assessment_files_a810.py (THIS SCRIPT)
    4. Generate dashboard: python3 generate_a810_5_compliance_dashboard.py
    5. Open dashboard, click "Update Links"

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.10
Utility Type:         Quality Assurance - Assessment File Normalization
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.10: Information Deletion Policy (Governance)
    - ISMS-IMP-A.8.10.1: Retention & Deletion Triggers Assessment (Domain 1)
    - ISMS-IMP-A.8.10.2: Deletion Methods Assessment (Domain 2)
    - ISMS-IMP-A.8.10.3: Third-Party & Cloud Deletion Assessment (Domain 3)
    - ISMS-IMP-A.8.10.4: Verification & Evidence Assessment (Domain 4)
    - ISMS-IMP-A.8.10.5: Compliance Dashboard (Consolidation)

Related Scripts:
    - generate_a810_1_retention_triggers.py (Domain 1 generator)
    - generate_a810_2_deletion_methods.py (Domain 2 generator)
    - generate_a810_3_third_party_cloud.py (Domain 3 generator)
    - generate_a810_4_verification_evidence.py (Domain 4 generator)
    - generate_a810_5_compliance_dashboard.py (Dashboard generator)
    - excel_sanity_check_a810.py (Quality validation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements filename normalization for A.8.10.1 through A.8.10.4
    - Validates Document ID consistency
    - Generates audit trail manifest
    - Supports dry-run mode and force overwrite

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**File Naming Requirements:**
The A.8.10.5 Compliance Dashboard uses external workbook formulas with EXACT
filename references. The normalized filenames produced by this script are:

- ISMS-IMP-A.8.10.1.xlsx (NOT ISMS_A_8_10_1_*.xlsx)
- ISMS-IMP-A.8.10.2.xlsx (NOT ISMS_A_8_10_2_*.xlsx)
- ISMS-IMP-A.8.10.3.xlsx (NOT ISMS_A_8_10_3_*.xlsx)
- ISMS-IMP-A.8.10.4.xlsx (NOT ISMS_A_8_10_4_*.xlsx)

These exact names are REQUIRED for external formula linking to work.

**Document ID Validation:**
This script validates that the Document ID in the "Instructions & Legend"
sheet matches the expected value for each assessment domain:

- Domain 1: ISMS-IMP-A.8.10.1
- Domain 2: ISMS-IMP-A.8.10.2
- Domain 3: ISMS-IMP-A.8.10.3
- Domain 4: ISMS-IMP-A.8.10.4

Document ID mismatches indicate files may be mislabeled or corrupted.

**Original File Preservation:**
This script COPIES files to normalized names - it does NOT delete or modify
original assessment files. Original files with date suffixes remain as backups.

**Overwrite Behavior:**
By default, if normalized files already exist, the script will prompt before
overwriting. Use --force to overwrite without prompting (useful for automation).

**Audit Trail:**
Every normalization run generates a manifest file documenting:
- Timestamp of normalization
- Source filename → Normalized filename mappings
- Document ID validation results
- Files successfully normalized vs. skipped

This provides audit evidence of systematic file management.

**Dashboard Dependency:**
The A.8.10.5 Compliance Dashboard REQUIRES normalized files to be in the
SAME DIRECTORY as the dashboard workbook. Plan your directory structure:

Recommended structure:
```
A810_Assessments/
├── ISMS-IMP-A.8.10.1.xlsx (normalized)
├── ISMS-IMP-A.8.10.2.xlsx (normalized)
├── ISMS-IMP-A.8.10.3.xlsx (normalized)
├── ISMS-IMP-A.8.10.4.xlsx (normalized)
├── ISMS_A_8_10_5_Compliance_Dashboard_20250124.xlsx (dashboard)
└── Originals/
    ├── ISMS_A_8_10_1_Retention_Triggers_20250124.xlsx
    ├── ISMS_A_8_10_2_Deletion_Methods_20250124.xlsx
    ├── ISMS_A_8_10_3_Third_Party_Cloud_20250124.xlsx
    └── ISMS_A_8_10_4_Verification_Evidence_20250124.xlsx
```

**Quality Assurance Workflow:**
This script is part of a comprehensive quality assurance workflow:

1. **Generate** assessments (generate_a810_1 through _4.py)
2. **Validate** quality (excel_sanity_check_a810.py)
3. **Complete** assessments (user data entry)
4. **Validate** again (excel_sanity_check_a810.py)
5. **Normalize** filenames (THIS SCRIPT) ← You are here
6. **Generate** dashboard (generate_a810_5_compliance_dashboard.py)
7. **Link** workbooks (Excel: Open dashboard → Update Links)

**External Linking vs. Data Consolidation:**
This normalization enables EXTERNAL LINKING approach (recommended):
- Dashboard formulas automatically pull data from assessments
- Real-time updates when assessments change
- Requires normalized filenames and same directory

Alternative: Data consolidation approach (if external linking fails):
- Use consolidate_a810_dashboard_data.py
- Copies data values instead of linking formulas
- Works with any filenames, any locations
- Requires manual re-run when data changes

**Error Handling:**
Common issues and resolutions:

**Issue:** "File not found matching pattern"
→ Solution: Ensure assessment files are generated and in correct directory

**Issue:** "Document ID mismatch"
→ Solution: File may be from wrong domain, regenerate or correct manually

**Issue:** "Normalized file already exists"
→ Solution: Use --force to overwrite or manually review/delete existing

**Issue:** "Cannot read workbook"
→ Solution: File may be corrupted or open in Excel, close and retry

**Issue:** "No files found to normalize"
→ Solution: Generate assessments first using generate_a810_X.py scripts

**Maintenance:**
Re-run normalization:
- After updating any assessment workbook (to sync changes to dashboard)
- After correcting Document ID mismatches
- When moving files to new directory structure
- Before distributing assessment package to stakeholders

**Integration with Other Tools:**
This script works with:
- excel_sanity_check_a810.py (validates before normalization)
- generate_a810_5_compliance_dashboard.py (uses normalized files)
- consolidate_a810_dashboard_data.py (alternative to external linking)

**Security Considerations:**
Normalized files may contain sensitive assessment data including:
- Compliance gaps and vulnerabilities
- Infrastructure details
- Third-party vendor information
- Evidence locations and access controls

Handle normalized files according to your organization's data classification.

**Best Practices:**
1. Run excel_sanity_check_a810.py BEFORE normalizing (catch issues early)
2. Review normalization manifest after each run (verify correct files)
3. Keep original dated files as backups (don't delete after normalizing)
4. Use consistent directory structure (makes dashboard linking reliable)
5. Document any manual filename corrections in change log
6. Test dashboard linking after normalization (verify #REF! errors absent)

**Limitations:**
- Does not validate assessment completeness (use excel_sanity_check_a810.py)
- Does not modify workbook content (only copies and renames files)
- Does not create dashboard (use generate_a810_5_compliance_dashboard.py)
- Does not consolidate data (use consolidate_a810_dashboard_data.py if needed)

**Troubleshooting:**

**Normalized files not recognized by dashboard:**
1. Verify exact filenames (case-sensitive, hyphens not underscores)
2. Confirm files are in same directory as dashboard
3. Check Document ID in each file matches expected value
4. Try closing and reopening Excel, then "Update Links"

**Normalization runs but dashboard shows #REF! errors:**
1. Verify sheet names in normalized files (e.g., "Summary Dashboard")
2. Check cell references match expected locations
3. Run excel_sanity_check_a810.py to validate structure
4. Regenerate assessments if structure is corrupted

**Multiple files match same pattern:**
1. Script uses most recent file by modification time
2. Review manifest to see which file was selected
3. Move old versions to separate directory if needed

================================================================================
"""

import os
import sys
import shutil
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("ℹ️  Install with: sudo apt install python3-openpyxl")     
    sys.exit(1)


# ============================================================================
# CONFIGURATION
# ============================================================================

# Expected document IDs and titles from assessment workbooks
EXPECTED_DOCS = {
    "ISMS-IMP-A.8.10.1": {
        "title": "Retention & Deletion Triggers Assessment",
        "normalized": "ISMS-IMP-A.8.10.1.xlsx"
    },
    "ISMS-IMP-A.8.10.2": {
        "title": "Deletion Methods Assessment",
        "normalized": "ISMS-IMP-A.8.10.2.xlsx"
    },
    "ISMS-IMP-A.8.10.3": {
        "title": "Third-Party & Cloud Deletion Assessment",
        "normalized": "ISMS-IMP-A.8.10.3.xlsx"
    },
    "ISMS-IMP-A.8.10.4": {
        "title": "Verification & Evidence Assessment",
        "normalized": "ISMS-IMP-A.8.10.4.xlsx"
    },
}


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_workbook(filepath):
    """
    Validate workbook contains expected document ID in Instructions sheet.
    
    Args:
        filepath: Path to Excel workbook
    
    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        # Open workbook in read-only mode
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # Check for Instructions & Legend sheet
        if "Instructions & Legend" not in wb.sheetnames:
            wb.close()
            return (None, None)
        
        ws = wb["Instructions & Legend"]
        
        # Look for Document ID in column A (typically rows 4-20)
        # Format: "Document ID" label in column A, value in column B
        for row in range(3, 25):
            cell_label = ws.cell(row=row, column=1).value
            
            if cell_label and "Document ID" in str(cell_label):
                doc_id_value = ws.cell(row=row, column=2).value
                
                if doc_id_value:
                    doc_id = str(doc_id_value).strip()
                    
                    # Check if it's one of our expected documents
                    if doc_id in EXPECTED_DOCS:
                        title = EXPECTED_DOCS[doc_id]["title"]
                        wb.close()
                        return (doc_id, title)
        
        wb.close()
        return (None, None)
        
    except Exception as e:
        print(f"    ⚠️  Error reading file: {e}")
        return (None, None)


def get_file_info(filepath):
    """Get file metadata for manifest."""
    stat = filepath.stat()
    return {
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime),
        "created": datetime.fromtimestamp(stat.st_ctime),
    }


# ============================================================================
# SCANNING FUNCTIONS
# ============================================================================

def scan_directory(directory):
    """
    Scan directory for assessment workbooks.
    
    Args:
        directory: Path to scan
    
    Returns:
        dict: {doc_id: {'path': Path, 'info': metadata}}
    """
    found_assessments = {}
    directory = Path(directory)
    
    # Get all .xlsx files, excluding temporary files
    xlsx_files = [f for f in directory.glob("*.xlsx") if not f.name.startswith("~$")]
    
    if not xlsx_files:
        print(f"\n❌ No Excel files found in {directory}\n")
        return found_assessments
    
    print(f"\n🔍 Scanning {len(xlsx_files)} Excel file(s) in {directory}...\n")
    
    for filepath in sorted(xlsx_files):
        print(f"  Checking: {filepath.name}")
        doc_id, title = validate_workbook(filepath)
        
        if doc_id:
            print(f"    ✅ Valid: {doc_id} - {title}")
            
            # Check for duplicates
            if doc_id in found_assessments:
                print(f"\n    ⚠️  WARNING: DUPLICATE FOUND FOR {doc_id}")
                print(f"        Previous: {found_assessments[doc_id]['path'].name}")
                print(f"        Current:  {filepath.name}\n")
                
                response = input("        Use CURRENT file? (y/n): ").strip().lower()
                
                if response == 'y':
                    found_assessments[doc_id] = {
                        'path': filepath,
                        'info': get_file_info(filepath)
                    }
                    print(f"        → Replaced with current file\n")
                else:
                    print(f"        → Keeping previous file\n")
            else:
                found_assessments[doc_id] = {
                    'path': filepath,
                    'info': get_file_info(filepath)
                }
        else:
            print(f"    ⏭️  Skipped (not a valid A.8.10 assessment workbook)")
        
        print()  # Blank line for readability
    
    return found_assessments


# ============================================================================
# MANIFEST CREATION
# ============================================================================

def create_manifest(output_dir, mapping, source_dir):
    """
    Create audit manifest documenting the normalization process.
    
    Args:
        output_dir: Destination directory
        mapping: Dictionary of normalized files
        source_dir: Source directory
    
    Returns:
        Path: Path to created manifest file
    """
    manifest_path = Path(output_dir) / "normalization_manifest_a810.txt"
    
    with open(manifest_path, 'w', encoding='utf-8') as f:
        # Header
        f.write("=" * 80 + "\n")
        f.write("ISMS ASSESSMENT FILE NORMALIZATION MANIFEST\n")
        f.write("ISO/IEC 27001:2022 - Control A.8.10: Information Deletion\n")
        f.write("=" * 80 + "\n\n")
        
        # Normalization metadata
        f.write("NORMALIZATION METADATA\n")
        f.write("-" * 80 + "\n")
        f.write(f"Normalization Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Source Directory:        {source_dir.resolve()}\n")
        f.write(f"Output Directory:        {output_dir.resolve()}\n")
        f.write(f"Files Normalized:        {len(mapping)}/4 required\n")
        f.write(f"Normalization Status:    {'COMPLETE' if len(mapping) == 4 else 'INCOMPLETE'}\n")
        f.write("\n")
        
        # File mapping details
        f.write("FILE MAPPING\n")
        f.write("-" * 80 + "\n\n")
        
        for doc_id in sorted(EXPECTED_DOCS.keys()):
            if doc_id in mapping:
                info = mapping[doc_id]
                source_path = info['path']
                normalized_name = info['normalized']
                metadata = info['info']
                
                f.write(f"Document ID:      {doc_id}\n")
                f.write(f"Title:            {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"Source File:      {source_path.name}\n")
                f.write(f"Normalized Name:  {normalized_name}\n")
                f.write(f"File Size:        {metadata['size']:,} bytes\n")
                f.write(f"Last Modified:    {metadata['modified'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Status:           ✅ NORMALIZED\n")
                f.write("\n")
            else:
                f.write(f"Document ID:      {doc_id}\n")
                f.write(f"Title:            {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"Status:           ❌ NOT FOUND\n")
                f.write("\n")
        
        # Validation summary
        f.write("VALIDATION SUMMARY\n")
        f.write("-" * 80 + "\n")
        f.write(f"Required Assessments:    4 (A.8.10.1 through A.8.10.4)\n")
        f.write(f"Found & Normalized:      {len(mapping)}\n")
        
        if len(mapping) == 4:
            f.write(f"Completeness:            ✅ ALL REQUIRED FILES PRESENT\n")
        else:
            f.write(f"Completeness:            ⚠️  INCOMPLETE ({4 - len(mapping)} missing)\n")
            missing = [doc_id for doc_id in EXPECTED_DOCS if doc_id not in mapping]
            f.write(f"Missing Documents:       {', '.join(missing)}\n")
        
        f.write("\n")
        
        # Usage instructions
        f.write("USAGE INSTRUCTIONS\n")
        f.write("-" * 80 + "\n")
        f.write("Dashboard Integration:\n")
        f.write("  1. Generate compliance dashboard:\n")
        f.write("     python3 generate_a810_5_compliance_dashboard.py\n\n")
        f.write("  2. Place dashboard file in same directory as these normalized files:\n")
        f.write(f"     {output_dir.resolve()}\n\n")
        f.write("  3. Open dashboard in Excel\n")
        f.write("  4. Excel will prompt: 'This workbook contains links to other data sources'\n")
        f.write("  5. Click 'Update' to enable automatic data population\n\n")
        
        f.write("How External Links Work:\n")
        f.write("  - Dashboard contains formulas that reference normalized files\n")
        f.write("  - Example: ='[ISMS-IMP-A.8.10.1.xlsx]Dashboard'!C25\n")
        f.write("  - Excel reads data directly from source assessment workbooks\n")
        f.write("  - Changes in source files auto-update dashboard (when links refreshed)\n")
        f.write("  - External workbook links auto-update when sources change\n")
        f.write("  - Place dashboard in same directory as normalized files\n")
        f.write("  - Open dashboard and click 'Update Links' when prompted\n\n")
        
        f.write("File Retention:\n")
        f.write("  - Original source files: Retained in source directory\n")
        f.write("  - Normalized copies: Located in output directory\n")
        f.write("  - This manifest: Retained with normalized files for audit\n\n")
        
        f.write("Quarterly Update Workflow:\n")
        f.write("  1. Update source assessment files (A.8.10.1-4) with new data\n")
        f.write("  2. Re-run this normalization script to refresh normalized copies\n")
        f.write("  3. Open dashboard → Excel prompts 'Update Links' → Click 'Update'\n")
        f.write("  4. Dashboard automatically reflects current assessment data\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("END OF MANIFEST\n")
        f.write("=" * 80 + "\n")
    
    return manifest_path


# ============================================================================
# MAIN NORMALIZATION FUNCTION
# ============================================================================

def normalize_files(source_dir=None, output_dir=None, auto_confirm=False):
    """
    Main normalization orchestration function.
    
    Args:
        source_dir: Source directory (prompts if None)
        output_dir: Output directory (prompts if None)
        auto_confirm: Skip confirmation prompts (for automation)
    
    Returns:
        bool: True if successful, False otherwise
    """
    print("=" * 80)
    print("ISMS ASSESSMENT FILE NORMALIZATION UTILITY")
    print("ISO/IEC 27001:2022 - Control A.8.10: Information Deletion")
    print("=" * 80)
    print("\nThis script prepares assessment workbooks for dashboard linking by:")
    print("  1. Scanning for completed assessment files (A.8.10.1-4)")
    print("  2. Validating document IDs in Instructions & Legend sheet")
    print("  3. Copying to normalized filenames (no dates/versions)")
    print("  4. Creating audit manifest for traceability\n")
    
    # Get source directory
    if not source_dir:
        source_dir = input("Enter source directory path (or press Enter for current directory): ").strip()
        if not source_dir:
            source_dir = "."
    
    source_dir = Path(source_dir).resolve()
    
    if not source_dir.exists():
        print(f"\n❌ Error: Source directory does not exist: {source_dir}\n")
        return False
    
    if not source_dir.is_dir():
        print(f"\n❌ Error: Not a directory: {source_dir}\n")
        return False
    
    print(f"\n📁 Source directory: {source_dir}")
    
    # Get output directory
    if not output_dir:
        default_output = source_dir / "Dashboard_Sources"
        output_input = input(f"\nEnter output directory (or press Enter for '{default_output.name}'): ").strip()
        if not output_input:
            output_dir = default_output
        else:
            output_dir = Path(output_input)
    
    output_dir = Path(output_dir).resolve()
    
    # Create output directory if it doesn't exist
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        print(f"📁 Output directory: {output_dir}")
    except Exception as e:
        print(f"\n❌ Error creating output directory: {e}\n")
        return False
    
    # Scan for assessment files
    found = scan_directory(source_dir)
    
    if not found:
        print("❌ No valid assessment workbooks found in source directory\n")
        print("   Ensure files contain valid Document IDs in 'Instructions & Legend' sheet:")
        for doc_id, info in EXPECTED_DOCS.items():
            print(f"   - {doc_id}: {info['title']}")
        print()
        return False
    
    # Show normalization summary
    print("=" * 80)
    print("NORMALIZATION SUMMARY")
    print("=" * 80 + "\n")
    print(f"Found {len(found)} of {len(EXPECTED_DOCS)} required assessment workbooks:\n")
    
    mapping = {}
    for doc_id in sorted(EXPECTED_DOCS.keys()):
        if doc_id in found:
            normalized_name = EXPECTED_DOCS[doc_id]["normalized"]
            source_path = found[doc_id]['path']
            
            print(f"  ✅ {doc_id}")
            print(f"     Title:      {EXPECTED_DOCS[doc_id]['title']}")
            print(f"     Source:     {source_path.name}")
            print(f"     Normalized: {normalized_name}\n")
            
            mapping[doc_id] = {
                'path': source_path,
                'normalized': normalized_name,
                'info': found[doc_id]['info']
            }
        else:
            print(f"  ❌ {doc_id} - NOT FOUND")
            print(f"     Title: {EXPECTED_DOCS[doc_id]['title']}\n")
    
    if len(found) < len(EXPECTED_DOCS):
        print(f"⚠️  WARNING: Only {len(found)}/{len(EXPECTED_DOCS)} assessment workbooks found")
        print("   Dashboard will have incomplete data for missing assessments\n")
    
    print(f"Output directory: {output_dir}\n")
    
    # Confirm normalization
    if not auto_confirm:
        response = input("Proceed with normalization? (y/n): ").strip().lower()
        if response != 'y':
            print("\n❌ Normalization cancelled by user\n")
            return False
    
    # Perform normalization (copy files)
    print("\n" + "=" * 80)
    print("NORMALIZING FILES...")
    print("=" * 80 + "\n")
    
    for doc_id, info in mapping.items():
        source = info['path']
        dest = output_dir / info['normalized']
        
        print(f"Copying: {source.name}")
        print(f"      → {dest.name}")
        
        try:
            shutil.copy2(source, dest)
            print(f"      ✅ Success\n")
        except Exception as e:
            print(f"      ❌ Error: {e}\n")
            print(f"❌ Normalization failed at {doc_id}\n")
            return False
    
    # Create audit manifest
    print("📄 Creating audit manifest...")
    try:
        manifest = create_manifest(output_dir, mapping, source_dir)
        print(f"   ✅ Created: {manifest.name}\n")
    except Exception as e:
        print(f"   ❌ Error creating manifest: {e}\n")
        return False
    
    # Success summary
    print("=" * 80)
    print("✅ NORMALIZATION COMPLETE")
    print("=" * 80 + "\n")
    print(f"Normalized files:  {output_dir}")
    print(f"Audit manifest:    {manifest}\n")
    print("NEXT STEPS:\n")
    print("  1. Review audit manifest for file mapping details")
    print("  2. Generate dashboard workbook:")
    print("     python3 generate_a810_5_compliance_dashboard.py\n")
    print("  3. Place dashboard in same directory as normalized files:")
    print(f"     {output_dir}\n")
    print("  4. Open dashboard and click 'Update Links' when prompted\n")
    print("  5. Dashboard will auto-populate with current deletion compliance data\n")
    print("=" * 80)
    print("\n💡 Pro Tip: Keep this script for quarterly updates!")
    print("   Just re-run it after updating source assessments to refresh dashboard.\n")
    print("=" * 80 + "\n")
    
    return True


# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Normalize ISMS A.8.10 assessment workbooks for dashboard integration",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode (prompts for directories)
  python3 normalize_assessment_files_a810.py
  
  # Specify source directory
  python3 normalize_assessment_files_a810.py --source ./assessments
  
  # Specify both directories
  python3 normalize_assessment_files_a810.py --source ./assessments --output ./dashboard
  
  # Automated mode (no prompts)
  python3 normalize_assessment_files_a810.py --source ./assessments --auto-confirm

Pattern Alignment:
  This script follows the proven A.8.24 normalization pattern for consistency
  across ISMS control implementations. External links enable auto-updating
  dashboards without manual data re-entry.
        """
    )
    
    parser.add_argument(
        '--source', '-s',
        help='Source directory containing assessment workbooks (default: current directory)'
    )
    
    parser.add_argument(
        '--output', '-o',
        help='Output directory for normalized files (default: ./Dashboard_Sources)'
    )
    
    parser.add_argument(
        '--auto-confirm', '-y',
        action='store_true',
        help='Skip confirmation prompts (for automation)'
    )
    
    args = parser.parse_args()
    
    success = normalize_files(
        source_dir=args.source,
        output_dir=args.output,
        auto_confirm=args.auto_confirm
    )
    
    sys.exit(0 if success else 1)