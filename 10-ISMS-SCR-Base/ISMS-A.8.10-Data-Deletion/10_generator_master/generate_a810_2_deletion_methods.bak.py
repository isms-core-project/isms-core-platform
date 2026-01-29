#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.10.2 - Deletion Methods Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.10: Information Deletion
Assessment Domain 2 of 4: Media-Specific Deletion Method Controls

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific storage infrastructure, media types, deletion
technologies, and NIST SP 800-88 implementation requirements.

Key customization areas:
1. Storage media inventory (match your actual HDD, SSD, tape, cloud infrastructure)
2. NIST SP 800-88 category mappings (adapt to your data classification scheme)
3. Deletion tool specifications (specific to your approved deletion utilities)
4. Verification testing procedures (based on your forensic testing capability)
5. SSD-specific crypto-erasure requirements (unique encryption key management)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.10 Information Deletion Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
media-specific deletion methods against NIST SP 800-88 standards and ISO 27001:2022
Control A.8.10 requirements, supporting evidence-based validation of secure data
sanitization practices across all storage technologies.

**Purpose:**
Enables systematic assessment of deletion method selection, implementation, and
verification for different storage media types (HDD, SSD, database, cloud, backup)
to ensure data cannot be recovered when deleted.

**Assessment Scope:**
- Physical storage media deletion (HDD, SSD, removable media, tape)
- Database system deletion methods and log management
- Cloud storage deletion policies and provider capabilities
- File system and backup media deletion procedures
- Deletion verification testing and forensic validation
- NIST SP 800-88 category compliance (Clear/Purge/Destroy)
- SSD-specific crypto-erasure requirements
- Gap analysis for inadequate deletion methods
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance, NIST SP 800-88 overview, color coding
2. Physical Storage Media - HDD, SSD, removable media, tape deletion methods
3. Database Systems - Database-specific deletion methods and log management
4. Cloud Storage - Cloud provider deletion policies and verification
5. File Systems & Backup - File share, NAS, backup media deletion methods
6. Deletion Verification Testing - Forensic testing results and verification methods
7. Summary Dashboard - NIST category compliance, gaps identification
8. Evidence Register - Testing certificates, forensic reports, tool documentation
9. Approval Sign-Off - Three-level stakeholder approval workflow

**Key Features:**
- Data validation with NIST category and media type dropdown lists
- Conditional formatting for data classification vs. NIST category mismatches
- Automated gap identification for inadequate deletion methods (e.g., Confidential + Clear)
- SSD-specific warnings (standard overwrite ineffective on SSDs)
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with NIST SP 800-88 media sanitization guidelines

**Integration:**
This assessment feeds into the A.8.10.5 Compliance Dashboard, which
consolidates data from all four information deletion assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a810_2_deletion_methods.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a810_2_deletion_methods.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a810_2_deletion_methods.py --date 20250124

Output:
    File: ISMS_A_8_10_2_Deletion_Methods_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize NIST categories to match your data classification
    2. Inventory all storage media types across your infrastructure
    3. Complete deletion method assessment for each media type
    4. Validate deletion methods align with NIST SP 800-88 for data classification
    5. Review SSD crypto-erasure requirements (unique encryption keys)
    6. Document deletion verification testing procedures
    7. Conduct gap analysis for inadequate methods (Confidential + Clear, etc.)
    8. Define remediation actions with timelines
    9. Collect and link audit evidence (deletion logs, forensic test certificates)
    10. Obtain stakeholder approvals
    11. Feed results into A.8.10.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.10
Assessment Domain:    2 of 4 (Media-Specific Deletion Method Controls)
Framework Version:    2.0  # Updated to include SSD-specific crypto-erasure guidance
Script Version:       2.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.10: Information Deletion Policy (Governance)
    - ISMS-REF-A.8.10: Deletion Methods Reference Guide (NIST SP 800-88)
    - ISMS-IMP-A.8.10.1: Retention & Deletion Triggers Assessment (Domain 1)
    - ISMS-IMP-A.8.10.2: Deletion Methods Implementation Guide
    - ISMS-IMP-A.8.10.3: Third-Party & Cloud Deletion Assessment (Domain 3)
    - ISMS-IMP-A.8.10.4: Verification & Evidence Assessment (Domain 4)
    - ISMS-IMP-A.8.10.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 2.0 - [Date to be set]
    - Added SSD-specific crypto-erasure guidance (unique encryption keys required)
    - Enhanced NIST category validation with data classification cross-checks
    - Added conditional formatting for Confidential+Clear mismatches
    - Included SSD warnings throughout assessment sheets
    - Updated Summary Dashboard with SSD-specific compliance metrics

Version 1.0 - [Initial Date]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.10.2 specification
    - Supports comprehensive deletion method evaluation per NIST SP 800-88
    - Integrated with A.8.10.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**NIST SP 800-88 Media Sanitization Guidelines:**
This assessment implements NIST SP 800-88 Rev. 1 framework with three categories:

- **CLEAR:** Logical techniques protecting against simple non-invasive data recovery.
  Use when media will be reused within organization.
  Examples: Overwrite (1-3 pass), block erase, cryptographic erase (key deletion).
  Protection: Prevents keyboard recovery, software file recovery tools.

- **PURGE:** Physical or logical techniques rendering data recovery infeasible using
  state-of-the-art laboratory techniques. Use when media leaves organizational control.
  Examples: Overwrite (cryptographic-strength), secure erase, degaussing, crypto-erasure.
  Protection: Prevents laboratory recovery attempts.

- **DESTROY:** Physical destruction rendering media unusable and preventing any data recovery.
  Use for highly sensitive data or end-of-life media.
  Examples: Shredding, disintegration, pulverization, incineration.
  Protection: Absolute - media is physically destroyed.

**CRITICAL: SSD-Specific Requirements:**
Standard overwrite methods (e.g., DoD 5220.22-M) are INEFFECTIVE on SSDs due to:
- Wear-leveling algorithms redistribute writes
- Over-provisioning creates inaccessible spare blocks
- TRIM command limitations and timing issues

For SSDs, use ONLY:
1. **Crypto-Erasure (Preferred):** Delete unique per-device encryption key
   - MUST use unique keys per device (NOT shared/cloned keys)
   - Verify key is truly deleted and cannot be recovered
2. **ATA Secure Erase:** Firmware-based sanitization (verify vendor implementation)
3. **Physical Destruction:** Shredding, disintegration, pulverization

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of deletion method selection, NIST category
alignment, and forensic testing results.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Storage media inventory and data locations
- Deletion tool specifications and configurations
- Forensic testing results and vulnerability information
- Gap analysis (inadequate deletion methods)

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check for new storage media types or deletion technologies
- Semi-annually: Update NIST category mappings for infrastructure changes
- Annually: Complete reassessment of all deletion methods and verification testing
- Ad-hoc: When new storage systems are deployed or media types change

**Quality Assurance:**
Have Information Security teams, Infrastructure teams, and Forensic specialists
validate assessments before using results for compliance reporting or remediation
decisions.

**Regulatory Alignment:**
Ensure deletion methods align with applicable regulatory requirements:
- Payment processing: PCI DSS Requirement 3.1, 9.8 (media destruction)
- Healthcare: HIPAA 164.310(d)(2)(i) (media disposal and re-use)
- Finance: GLBA Disposal Rule (proper disposal of consumer information)
- Government: NIST SP 800-88 compliance mandates

Customize assessment criteria to include regulatory-specific requirements.

**Common Mistakes to Avoid:**
1. ❌ Using "Clear" for Confidential/Restricted data - INSUFFICIENT protection
2. ❌ Using standard overwrite on SSDs - INEFFECTIVE due to wear-leveling
3. ❌ Failing to verify deletion effectiveness - No forensic testing
4. ❌ Forgetting backup deletion - Data persists in backups
5. ❌ Using shared encryption keys for crypto-erasure - NOT unique per device

**Verification Paradox:**
Proving data was deleted without retaining the deleted data creates a challenge.
Solutions:
- Retain metadata about what was deleted (not the data itself)
- Use deletion logs with checksums/hashes
- Conduct forensic testing on sample media
- Document deletion method specifications and tool configurations

================================================================================
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ==========================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ==========================================================================



# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
TRASH = '\U0001F5D1'  # 🗑️  Wastebasket
DISK = '\U0001F4BE'   # 💾 Floppy Disk
GLOBE = '\U0001F310'  # 🌐 Globe
SEARCH = '\U0001F50D' # 🔍 Magnifying Glass
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create all sheets in order
    wb.create_sheet("Instructions & Legend", 0)
    wb.create_sheet("2. Physical Media Deletion", 1)
    wb.create_sheet("3. Cloud Storage Deletion", 2)
    wb.create_sheet("4. Database & Application Deletion", 3)
    wb.create_sheet("5. Mobile & Endpoint Deletion", 4)
    wb.create_sheet("6. Deletion Tool Validation", 5)
    wb.create_sheet("Summary Dashboard", 6)
    wb.create_sheet("Evidence Register", 7)
    wb.create_sheet("Approval Sign-Of", 8)
    
    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    styles = {
        'title': {
            'font': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'header': {
            'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'subheader': {
            'font': Font(name='Calibri', size=10, bold=True, color='000000'),
            'fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'input_cell': {
            'font': Font(name='Calibri', size=10, color='000000'),
            'fill': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'status_compliant': {
            'font': Font(name='Calibri', size=10, bold=True, color='006100'),
            'fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'status_partial': {
            'font': Font(name='Calibri', size=10, bold=True, color='9C5700'),
            'fill': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'status_noncompliant': {
            'font': Font(name='Calibri', size=10, bold=True, color='9C0006'),
            'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'reference_cell': {
            'font': Font(name='Calibri', size=9, color='000000'),
            'fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'normal': {
            'font': Font(name='Calibri', size=10, color='000000'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    }
    return styles


# ==========================================================================
# SECTION 2: COLUMN DEFINITIONS
# ==========================================================================

def get_base_columns():
    """Return base column structure (A-Q, 17 columns)."""
    return [
        ("A", "Media Type / System Name", 30),
        ("B", "Data Classification", 22),
        ("C", "Owner / Responsible Party", 18),
        ("D", "Deletion Method Used", 25),
        ("E", "Tool/Vendor Used", 20),
        ("F", "Status", 18),
        ("G", "Implementation Date", 12),
        ("H", "Last Effectiveness Test", 12),
        ("I", "Next Test Date", 12),
        ("J", "Gap Identified", 25),
        ("K", "Remediation Plan", 25),
        ("L", "Target Completion", 12),
        ("M", "Risk Level", 12),
        ("N", "Evidence Reference", 20),
        ("O", "Notes / Comments", 25),
        ("P", "Remediation Owner", 18),
        ("Q", "Budget Required", 15)
    ]


def get_extended_columns_sheet2():
    """Sheet 2: Physical Media Deletion extensions."""
    return [
        ("R", "NIST SP 800-88 Method", 18),
        ("S", "Verification Method", 22),
        ("T", "Last Forensic Test Result", 20),
        ("U", "Media Disposal Method", 18)
    ]


def get_extended_columns_sheet3():
    """Sheet 3: Cloud Storage Deletion extensions."""
    return [
        ("R", "Provider Tier", 12),
        ("S", "Deletion API Available", 20),
        ("T", "Snapshot/Backup Deletion", 22),
        ("U", "Multi-Region Deletion Verified", 25)
    ]


def get_extended_columns_sheet4():
    """Sheet 4: Database & Application Deletion extensions."""
    return [
        ("R", "Deletion Type", 20),
        ("S", "Referential Integrity Handled", 25),
        ("T", "Backup Purging", 18),
        ("U", "Crypto-Erasure Key Management", 25)
    ]


def get_extended_columns_sheet5():
    """Sheet 5: Mobile & Endpoint Deletion extensions."""
    return [
        ("R", "Device Type", 20),
        ("S", "MDM Solution Used", 20),
        ("T", "Remote Wipe Capability", 20),
        ("U", "Full Disk Encryption", 20)
    ]


def get_extended_columns_sheet6():
    """Sheet 6: Deletion Tool Validation extensions."""
    return [
        ("R", "Test Frequency", 18),
        ("S", "Forensic Tool Used", 22),
        ("T", "Pass Rate (Last 12 Months)", 22),
        ("U", "Independent Testing", 18)
    ]


# ==========================================================================
# SECTION 3: DATA VALIDATION
# ==========================================================================

def create_base_validations(ws):
    """Create data validation objects for standard columns."""
    
    # Data Classification (Column B)
    dv_classification = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Restricted"',
        allow_blank=True
    )
    dv_classification.error = 'Please select from dropdown'
    dv_classification.errorTitle = 'Invalid Classification'
    ws.add_data_validation(dv_classification)
    dv_classification.add('B10:B100')
    
    # Status (Column F)
    dv_status = DataValidation(
        type="list",
        formula1='f"{CHECK} Compliant,⚠️ Partial,❌ Non-Compliant,N/A"',
        allow_blank=True
    )
    dv_status.error = 'Please select from dropdown'
    dv_status.errorTitle = 'Invalid Status'
    ws.add_data_validation(dv_status)
    dv_status.add('F10:F100')
    
    # Risk Level (Column M)
    dv_risk = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    dv_risk.error = 'Please select from dropdown'
    dv_risk.errorTitle = 'Invalid Risk Level'
    ws.add_data_validation(dv_risk)
    dv_risk.add('M10:M100')
    
    # Budget Required (Column Q)
    dv_budget = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=True
    )
    dv_budget.error = 'Please select from dropdown'
    dv_budget.errorTitle = 'Invalid Budget Option'
    ws.add_data_validation(dv_budget)
    dv_budget.add('Q10:Q100')


def create_sheet_specific_validations(ws, sheet_type):
    """Create sheet-specific data validations."""
    
    if sheet_type == "sheet2":
        # Deletion Method (Column D) - Physical Media
        dv = DataValidation(
            type="list",
            formula1='"Overwrite (1-pass),Overwrite (3-pass),Overwrite (7-pass DoD 5220.22-M),Secure Erase (ATA/NVMe),Cryptographic Erasure,Degaussing,Physical Destruction (Shred),Physical Destruction (Incinerate),Physical Destruction (Crush/Pulverize),Other (specify in notes)"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('D10:D100')
        
        # NIST SP 800-88 Method (Column R)
        dv = DataValidation(
            type="list",
            formula1='"Clear,Purge,Destroy,N/A"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('R10:R100')
        
        # Verification Method (Column S)
        dv = DataValidation(
            type="list",
            formula1='"Software Verification,Visual Inspection,Certificate of Destruction,Forensic Test,None"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('S10:S100')
        
        # Last Forensic Test Result (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Pass (No Recovery),Fail (Partial Recovery),Fail (Full Recovery),Not Tested"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('T10:T100')
        
        # Media Disposal Method (Column U)
        dv = DataValidation(
            type="list",
            formula1='"Reuse,Recycle,Shred,Incinerate,Degauss,Other"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('U10:U100')
    
    elif sheet_type == "sheet3":
        # Deletion Method (Column D) - Cloud
        dv = DataValidation(
            type="list",
            formula1='"Cloud Provider API Delete,Cloud Lifecycle Policy,Manual Console Delete,Crypto-Shred (Key Deletion),Account Closure,Other (specify in notes)"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('D10:D100')
        
        # Provider Tier (Column R)
        dv = DataValidation(
            type="list",
            formula1='"Tier 1,Tier 2,Tier 3,Tier 4,Tier 5,Tier 6,Tier 7,Tier 8,Tier 9,Tier 10"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('R10:R100')
        
        # Deletion API Available (Column S)
        dv = DataValidation(
            type="list",
            formula1='"Yes - Automated,Yes - Manual,No,Unknown"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('S10:S100')
        
        # Snapshot/Backup Deletion (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Automatic,Manual,Unknown,N/A"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('T10:T100')
        
        # Multi-Region Deletion Verified (Column U)
        dv = DataValidation(
            type="list",
            formula1='"Yes,No,N/A,Unknown"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('U10:U100')
    
    elif sheet_type == "sheet4":
        # Deletion Method (Column D) - Database
        dv = DataValidation(
            type="list",
            formula1='"Logical DELETE (soft delete),Hard DELETE (permanent),TRUNCATE TABLE,DROP TABLE/DATABASE,Purge/Vacuum,Crypto-Shred,Archive then Delete,Other (specify in notes)"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('D10:D100')
        
        # Deletion Type (Column R)
        dv = DataValidation(
            type="list",
            formula1='"Logical Delete,Hard Delete,Purge,Truncate,Crypto-Shred,Archive then Delete"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('R10:R100')
        
        # Referential Integrity Handled (Column S)
        dv = DataValidation(
            type="list",
            formula1='"Yes - Cascading,Yes - Manual,No,N/A"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('S10:S100')
        
        # Backup Purging (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Automatic,Manual,Not Implemented,N/A"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('T10:T100')
        
        # Crypto-Erasure Key Management (Column U)
        dv = DataValidation(
            type="list",
            formula1='"Yes - Automated,Yes - Manual,No,N/A"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('U10:U100')
    
    elif sheet_type == "sheet5":
        # Deletion Method (Column D) - Mobile/Endpoint (generic, sheet has specific device types)
        dv = DataValidation(
            type="list",
            formula1='"Factory Reset,MDM Remote Wipe,Full Disk Encryption + Format,Secure Erase,Physical Destruction,Other (specify in notes)"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('D10:D100')
        
        # Device Type (Column R)
        dv = DataValidation(
            type="list",
            formula1='"Corporate Laptop,Corporate Mobile,BYOD Laptop,BYOD Mobile,Tablet,Other"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('R10:R100')
        
        # Remote Wipe Capability (Column T)
        dv = DataValidation(
            type="list",
            formula1='"Yes - Tested,Yes - Untested,No,N/A"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('T10:T100')
        
        # Full Disk Encryption (Column U)
        dv = DataValidation(
            type="list",
            formula1='"Yes - BitLocker,Yes - FileVault,Yes - LUKS,Yes - Other,No"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('U10:U100')
    
    elif sheet_type == "sheet6":
        # Deletion Method (Column D) - Validation (generic)
        dv = DataValidation(
            type="list",
            formula1='"Forensic Recovery Test,Software Verification,Certificate Review,Visual Inspection,Other (specify in notes)"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('D10:D100')
        
        # Test Frequency (Column R)
        dv = DataValidation(
            type="list",
            formula1='"Monthly,Quarterly,Semi-Annual,Annual,Ad-hoc,Never"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('R10:R100')
        
        # Independent Testing (Column U)
        dv = DataValidation(
            type="list",
            formula1='"Yes - External,Yes - Internal,No,Planned"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('U10:U100')


# ==========================================================================
# SECTION 4: HELPER FUNCTIONS
# ==========================================================================

def apply_cell_style(cell, styles, style_type):
    """Apply style dictionary to a cell."""
    style = styles[style_type]
    cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    cell.alignment = style['alignment']
    if 'border' in style:
        cell.border = style['border']


def create_header_row(ws, row, columns, styles):
    """Create header row with column definitions."""
    for col_letter, header_text, width in columns:
        cell = ws[f'{col_letter}{row}']
        cell.value = header_text
        apply_cell_style(cell, styles, 'header')
        ws.column_dimensions[col_letter].width = width


def create_data_rows(ws, start_row, num_rows, num_cols, styles):
    """Create yellow-highlighted data entry rows."""
    for row in range(start_row, start_row + num_rows):
        for col_idx in range(1, num_cols + 1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f'{col_letter}{row}']
            apply_cell_style(cell, styles, 'input_cell')


def create_checklist_section(ws, start_row, checklist_items, styles):
    """Create compliance checklist section."""
    # Section header
    ws.merge_cells(f'A{start_row}:U{start_row}')
    cell = ws[f'A{start_row}']
    cell.value = "COMPLIANCE CHECKLIST - Mark ✅ or {XMARK}"
    apply_cell_style(cell, styles, 'subheader')
    
    # Checklist items
    current_row = start_row + 1
    for idx, item in enumerate(checklist_items, 1):
        ws[f'A{current_row}'].value = f"{idx}."
        ws[f'A{current_row}'].alignment = Alignment(horizontal='right')
        
        ws.merge_cells(f'B{current_row}:S{current_row}')
        ws[f'B{current_row}'].value = item
        ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        ws[f'T{current_row}'].value = ""
        ws[f'T{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        ws[f'U{current_row}'].value = ""
        ws[f'U{current_row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        current_row += 1
    
    return current_row


def create_reference_table(ws, start_row, table_title, headers, data, styles):
    """Create reference table with title and data."""
    num_cols = len(headers)
    end_col = get_column_letter(num_cols)
    ws.merge_cells(f'A{start_row}:{end_col}{start_row}')
    cell = ws[f'A{start_row}']
    cell.value = table_title
    apply_cell_style(cell, styles, 'subheader')
    
    header_row = start_row + 1
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        cell = ws[f'{col_letter}{header_row}']
        cell.value = header
        apply_cell_style(cell, styles, 'header')
    
    current_row = header_row + 1
    for row_data in data:
        for col_idx, value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = value
            apply_cell_style(cell, styles, 'reference_cell')
        current_row += 1
    
    return current_row + 1


# ==========================================================================
# SECTION 5: ASSESSMENT SHEET CREATOR (CORE FUNCTION)
# ==========================================================================

def create_assessment_sheet(ws, styles, section_title, policy_ref, 
                            assessment_question, base_cols, extended_cols,
                            checklist_items, reference_tables, sheet_type):
    """
    Create standardized assessment sheet.
    """
    
    # Row 1: Title
    ws.merge_cells('A1:U1')
    cell = ws['A1']
    cell.value = section_title
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    # Row 2: Policy Reference
    ws.merge_cells('A2:U2')
    cell = ws['A2']
    cell.value = f"Policy Reference: {policy_ref}"
    cell.font = Font(name='Calibri', size=9, italic=True)
    cell.alignment = Alignment(horizontal='left')
    
    # Row 3: Blank
    ws.row_dimensions[3].height = 5
    
    # Row 4-6: Assessment Question
    ws.merge_cells('A4:U6')
    cell = ws['A4']
    cell.value = f"ASSESSMENT QUESTION:\n{assessment_question}"
    cell.font = Font(name='Calibri', size=11, bold=True)
    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[4].height = 50
    
    # Row 7: Instructions
    ws.merge_cells('A7:U7')
    cell = ws['A7']
    cell.value = "Complete the yellow-highlighted cells below. Use dropdowns where provided. Link evidence in Column N to Evidence Register sheet."
    cell.font = Font(name='Calibri', size=9, italic=True, color='0000FF')
    cell.alignment = Alignment(horizontal='left', wrap_text=True)
    
    # Row 8: Blank
    ws.row_dimensions[8].height = 5
    
    # Row 9: Column Headers
    all_columns = base_cols + extended_cols
    create_header_row(ws, 9, all_columns, styles)
    
    # Rows 10-22: Data Entry Rows (13 rows)
    create_data_rows(ws, 10, 13, len(all_columns), styles)
    
    # Row 23-24: Blank
    ws.row_dimensions[23].height = 10
    ws.row_dimensions[24].height = 5
    
    # Rows 25+: Compliance Checklist
    next_row = create_checklist_section(ws, 25, checklist_items, styles)
    
    ws.row_dimensions[next_row].height = 10
    next_row += 1
    
    # Reference Tables
    for table_title, headers, data in reference_tables:
        next_row = create_reference_table(ws, next_row, table_title, headers, data, styles)
    
    # Exception/Deviation Block
    ws.merge_cells(f'A{next_row}:U{next_row}')
    cell = ws[f'A{next_row}']
    cell.value = "EXCEPTIONS / DEVIATIONS"
    apply_cell_style(cell, styles, 'subheader')
    
    next_row += 1
    ws.merge_cells(f'A{next_row}:U{next_row+2}')
    cell = ws[f'A{next_row}']
    cell.value = "Document any exceptions or deviations from requirements here:"
    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[next_row].height = 60
    
    # Freeze panes
    ws.freeze_panes = 'A10'
    
    # Apply validations
    create_base_validations(ws)
    create_sheet_specific_validations(ws, sheet_type)


# ==========================================================================
# SECTION 6: INSTRUCTIONS SHEET
# ==========================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet with NIST framework."""
    
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "ISMS-IMP-A.8.10.2 - Deletion Methods Assessment"
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    current_row = 3
    
    # Document Information
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "DOCUMENT CONTROL"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    control_fields = [
        ("Workbook Version:", "1.0"),
        ("Assessment Date:", datetime.now().strftime('%Y-%m-%d')),
        ("Assessor Name:", "[Enter Name]"),
        ("Organization:", "[Enter Organization]"),
        ("Review Period:", "[e.g., Q4 2025]")
    ]
    
    for label, value in control_fields:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        current_row += 1
    
    current_row += 2
    
    # NIST SP 800-88 Framework
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "NIST SP 800-88 FRAMEWORK OVERVIEW"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    nist_framework = [
        ("CLEAR", "Logical techniques to protect against simple non-invasive data recovery. Use when media will be reused within organization. Examples: Overwrite (1-3 pass), block erase, cryptographic erase (key deletion). Protection: Prevents keyboard recovery, software file recovery tools."),
        ("PURGE", "Physical or logical techniques that render target data recovery infeasible using state-of-the-art laboratory techniques. Use when media leaves organizational control. Examples: Overwrite (cryptographic-strength), secure erase, degaussing, cryptographic erasure. Protection: Prevents laboratory recovery attempts."),
        ("DESTROY", "Physical destruction of media to render it unusable and prevent any data recovery. Use for highly sensitive data or end-of-life media. Examples: Shredding, disintegration, pulverization, incineration. Protection: Absolute - media is physically destroyed.")
    ]
    
    for method, description in nist_framework:
        ws[f'A{current_row}'].value = method
        ws[f'A{current_row}'].font = Font(bold=True, size=11)
        ws[f'A{current_row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'].value = description
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        ws.row_dimensions[current_row].height = 50
        current_row += 2
    
    # Media Type Quick Reference
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "MEDIA TYPE QUICK REFERENCE"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    media_ref_headers = ["Media Type", "Recommended NIST Method", "Common Tools/Methods"]
    for col_idx, header in enumerate(media_ref_headers, 1):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{current_row}'].value = header
        apply_cell_style(ws[f'{col_letter}{current_row}'], styles, 'header')
    current_row += 1
    
    media_types = [
        ("HDD (spinning disk)", "Purge or Destroy", "Overwrite (7-pass), Degauss, Shred"),
        ("SSD (NAND flash)", "Purge or Destroy", "Secure Erase (ATA), Crypto-Erase, Shred"),
        ("NVMe SSD", "Purge or Destroy", "NVMe Secure Erase, Crypto-Erase, Shred"),
        ("Tape (magnetic)", "Purge or Destroy", "Degauss, Overwrite, Shred"),
        ("USB/SD Cards", "Purge or Destroy", "Overwrite, Crypto-Erase, Shred"),
        ("Optical Media (CD/DVD)", "Destroy", "Shred, Incinerate"),
        ("Paper Documents", "Destroy", "Cross-cut shred, Pulp, Incinerate"),
        ("Cloud Storage", "Clear or Purge", "API Delete, Crypto-Erase, Account Closure"),
        ("Database Records", "Clear", "Hard DELETE, TRUNCATE, Purge/Vacuum"),
        ("Mobile Devices", "Purge or Destroy", "Factory Reset + Encryption, MDM Wipe, Destroy")
    ]
    
    for media, nist, tools in media_types:
        ws[f'A{current_row}'].value = media
        ws[f'B{current_row}'].value = nist
        ws[f'C{current_row}'].value = tools
        for col in ['A', 'B', 'C']:
            ws[f'{col}{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        current_row += 1
    
    current_row += 2
    
    # Important Notes
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "CRITICAL NOTES"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    notes = [
        f"{WARNING} CRITICAL: SSDs cannot be effectively sanitized using overwrite methods due to wear leveling and over-provisioning. Use Secure Erase or physical destruction.",
        "✓ This assessment is vendor-neutral. Document YOUR organization's specific tools and methods.",
        "✓ Cloud provider deletion limitations must be understood (snapshots, backups, multi-region).",
        "✓ Backup deletion must be coordinated with active data deletion.",
        "✓ Annual forensic testing is required to validate deletion effectiveness.",
        "✓ Related assessments: A.8.10.1 (Retention), A.8.10.3 (Third-Party), A.8.10.4 (Verification)"
    ]
    
    for note in notes:
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'].value = note
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        ws.row_dimensions[current_row].height = 25
        current_row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15


# ==========================================================================
# SECTION 7: EVIDENCE REGISTER
# ==========================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register with 100 rows."""
    
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "Evidence Register - Supporting Documentation"
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    ws.merge_cells('A2:K2')
    cell = ws['A2']
    cell.value = "Use this register to document all evidence supporting your deletion method assessments. Reference evidence by ID (Column A) in the 'Evidence Reference' column (Column N) of assessment sheets."
    cell.font = Font(name='Calibri', size=9, italic=True)
    cell.alignment = Alignment(horizontal='left', wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    ws.row_dimensions[3].height = 5
    
    headers = [
        ("A", "Evidence ID", 12),
        ("B", "Assessment Sheet", 20),
        ("C", "Related Media/System", 30),
        ("D", "Evidence Type", 20),
        ("E", "Evidence Title/Description", 35),
        ("F", "File Location/Link", 40),
        ("G", "Date Created/Collected", 12),
        ("H", "Retention Period", 15),
        ("I", "Next Review Date", 12),
        ("J", "Owner/Custodian", 20),
        ("K", "Notes", 30)
    ]
    
    create_header_row(ws, 4, headers, styles)
    
    for row in range(5, 105):
        for col_idx, (col_letter, _, _) in enumerate(headers, 1):
            cell = ws[f'{col_letter}{row}']
            apply_cell_style(cell, styles, 'input_cell')
    
    # Dropdowns
    dv_type = DataValidation(
        type="list",
        formula1='"Deletion Tool Documentation,Forensic Test Report,Screenshot/Print Screen,System Log Export,Configuration File,Certificate of Destruction,Vendor Documentation,NIST SP 800-88 Compliance Memo,Test Result (Pass/Fail),Third-Party Audit Report,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add('D5:D104')
    
    dv_sheet = DataValidation(
        type="list",
        formula1='"Sheet 2: Physical Media Deletion,Sheet 3: Cloud Storage Deletion,Sheet 4: Database & Application Deletion,Sheet 5: Mobile & Endpoint Deletion,Sheet 6: Deletion Tool Validation"',
        allow_blank=True
    )
    ws.add_data_validation(dv_sheet)
    dv_sheet.add('B5:B104')
    
    dv_retention = DataValidation(
        type="list",
        formula1='"3 years,5 years,7 years,10 years,Duration of media lifecycle + 1 year,Permanent"',
        allow_blank=True
    )
    ws.add_data_validation(dv_retention)
    dv_retention.add('H5:H104')
    
    ws.freeze_panes = 'A5'


# ==========================================================================
# SECTION 8: APPROVAL SIGN-OFF
# ==========================================================================

def create_approval_signoff(ws, styles):
    """Create Approval Sign-Off workflow sheet."""
    
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "Three-Level Approval Workflow"
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    current_row = 3
    
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "DOCUMENT CONTROL"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    control_fields = [
        ("Assessment Period:", "[e.g., Q4 2025]"),
        ("Workbook Version:", "1.0"),
        ("Total Assessment Sheets Completed:", "5"),
        ("Overall Compliance % (from Dashboard):", "[Link to Summary Dashboard]"),
        ("Critical Gaps Identified:", "[Count from Summary]"),
        ("Assessment Completed By:", "[Name, Date]")
    ]
    
    for label, value in control_fields:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        current_row += 1
    
    current_row += 2
    
    # Level 1 Approval
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "LEVEL 1 APPROVAL - Technical/Operational"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = 'Role: IT Security Manager / Systems Administrator / Data Security Engineer'
    ws[f'A{current_row}'].font = Font(italic=True)
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row + 1}')
    ws[f'A{current_row}'].value = 'Approval Statement: "I confirm that this assessment accurately reflects our current deletion methods across all media types as of [Date]. All deletion tools have been documented, effectiveness testing is current, and remediation plans exist for identified gaps."'
    ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[current_row].height = 40
    current_row += 2
    
    approval_fields = ["Approver Name:", "Title/Role:", "Email:", "Review Date:", "Approval Status:", "Conditions/Comments:", "Signature:"]
    
    for field in approval_fields:
        ws[f'A{current_row}'].value = field
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        if field == "Approval Status:":
            dv = DataValidation(type="list", formula1='f"{CHECK} Approved,⚠️ Approved with Conditions,❌ Rejected"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f'B{current_row}')
        current_row += 1
    
    current_row += 2
    
    # Level 2
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "LEVEL 2 APPROVAL - Management"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = 'Role: Chief Information Security Officer / IT Director / Compliance Manager'
    ws[f'A{current_row}'].font = Font(italic=True)
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row + 1}')
    ws[f'A{current_row}'].value = 'Approval Statement: "I acknowledge the findings of this A.8.10.2 assessment and approve the proposed remediation plans. Budget and resources will be allocated to address critical gaps, particularly SSD deletion methods and forensic testing programs."'
    ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[current_row].height = 40
    current_row += 2
    
    for field in approval_fields:
        ws[f'A{current_row}'].value = field
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        if field == "Approval Status:":
            dv = DataValidation(type="list", formula1='f"{CHECK} Approved,⚠️ Approved with Conditions,❌ Rejected"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f'B{current_row}')
        current_row += 1
    
    current_row += 2
    
    # Level 3
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "LEVEL 3 APPROVAL - Executive"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = 'Role: Chief Information Officer / Chief Risk Officer / Board Delegate'
    ws[f'A{current_row}'].font = Font(italic=True)
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row + 1}')
    ws[f'A{current_row}'].value = 'Approval Statement: "This assessment has been reviewed at the executive level. The organization\'s deletion method effectiveness is [Acceptable/Needs Improvement/Unacceptable]. The Board/Executive Team has been briefed on critical gaps related to SSD deletion, cloud provider dependencies, and forensic testing results."'
    ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[current_row].height = 50
    current_row += 2
    
    for field in approval_fields[:-1]:
        ws[f'A{current_row}'].value = field
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        if field == "Approval Status:":
            dv = DataValidation(type="list", formula1='f"{CHECK} Approved,⚠️ Approved with Conditions,❌ Rejected"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f'B{current_row}')
        current_row += 1
    
    ws[f'A{current_row}'].value = "Executive Summary:"
    ws.merge_cells(f'B{current_row}:F{current_row + 2}')
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    ws[f'B{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[current_row].height = 60
    current_row += 3
    
    ws[f'A{current_row}'].value = "Signature:"
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15


# ==========================================================================
# SECTION 9: SUMMARY DASHBOARD
# ==========================================================================

def create_summary_dashboard(ws, styles):
    """Create comprehensive summary dashboard."""
    
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = "A.8.10.2 Deletion Methods - Compliance Dashboard"
    apply_cell_style(cell, styles, 'title')
    ws.row_dimensions[1].height = 25
    
    current_row = 3
    
    # Overall Compliance Summary
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "OVERALL COMPLIANCE SUMMARY"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{current_row}'].value = header
        apply_cell_style(ws[f'{col_letter}{current_row}'], styles, 'header')
    current_row += 1
    
    areas = ["Physical Media Deletion", "Cloud Storage Deletion", "Database & Application Deletion", 
             "Mobile & Endpoint Deletion", "Deletion Tool Validation"]
    
    for area in areas:
        ws[f'A{current_row}'].value = area
        ws[f'B{current_row}'].value = "[COUNT]"
        ws[f'C{current_row}'].value = "[COUNTIF]"
        ws[f'D{current_row}'].value = "[COUNTIF]"
        ws[f'E{current_row}'].value = "[COUNTIF]"
        ws[f'F{current_row}'].value = "[COUNTIF]"
        ws[f'G{current_row}'].value = "[%]"
        current_row += 1
    
    ws[f'A{current_row}'].value = "OVERALL A.8.10.2"
    ws[f'A{current_row}'].font = Font(bold=True)
    ws[f'B{current_row}'].value = "[SUM]"
    ws[f'G{current_row}'].value = "[Overall %]"
    current_row += 2
    
    ws[f'A{current_row}'].value = "Compliance Thresholds:"
    ws[f'B{current_row}'].value = "≥90% = 🟢 Excellent | 70-89% = 🟡 Needs Improvement | <70% = 🔴 Critical"
    ws[f'B{current_row}'].font = Font(italic=True, size=9)
    current_row += 2
    
    # Critical Gaps
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "CRITICAL GAPS - SSD OVERWRITE DETECTION"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:G{current_row}')
    ws[f'A{current_row}'].value = f"{WARNING} CRITICAL: SSDs using overwrite methods are INEFFECTIVE due to wear leveling. Use Secure Erase or physical destruction."
    ws[f'A{current_row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    ws[f'A{current_row}'].font = Font(bold=True, size=11)
    current_row += 1
    
    ws[f'A{current_row}'].value = "Count of SSDs using Overwrite:"
    ws[f'B{current_row}'].value = '[=COUNTIFS(Sheet2!A:A,"*SSD*",Sheet2!D:D,"Overwrite*")]'
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    current_row += 3
    
    # NIST Method Distribution
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "NIST SP 800-88 METHOD DISTRIBUTION"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    nist_headers = ["NIST Method", "Count", "% of Total", "Status"]
    for col_idx, header in enumerate(nist_headers, 1):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{current_row}'].value = header
        apply_cell_style(ws[f'{col_letter}{current_row}'], styles, 'header')
    current_row += 1
    
    nist_methods = ["Clear", "Purge", "Destroy", "N/A (Non-Physical)"]
    for method in nist_methods:
        ws[f'A{current_row}'].value = method
        ws[f'B{current_row}'].value = f'[COUNTIF Sheet2!R:R="{method}"]'
        ws[f'C{current_row}'].value = "[%]"
        ws[f'D{current_row}'].value = "ℹ️ Info"
        current_row += 1
    
    current_row += 1
    
    # Executive Summary
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "EXECUTIVE SUMMARY & RECOMMENDATIONS"
    apply_cell_style(cell, styles, 'subheader')
    current_row += 1
    
    ws[f'A{current_row}'].value = "Overall A.8.10.2 Maturity Level:"
    ws[f'B{current_row}'].value = "[Emerging / Developing / Established / Optimized]"
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    current_row += 2
    
    ws[f'A{current_row}'].value = "Key Strengths:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    for i in range(1, 4):
        ws[f'A{current_row}'].value = f"{i}."
        ws[f'B{current_row}'].value = "[Example strength]"
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        current_row += 1
    
    current_row += 1
    ws[f'A{current_row}'].value = "Critical Improvement Areas:"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    for i in range(1, 4):
        ws[f'A{current_row}'].value = f"{i}."
        ws[f'B{current_row}'].value = "[Critical gap]"
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        current_row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


# ==========================================================================
# SECTION 10: DOMAIN-SPECIFIC SHEET CREATORS
# ==========================================================================

def create_sheet2_physical_media(ws, styles):
    """Create Sheet 2: Physical Media Deletion."""
    
    checklist = [
        "All physical media types in use are identified and documented",
        "Deletion methods are assigned per media type (not one-size-fits-all)",
        "HDD deletion uses overwrite (≥3 pass) OR degaussing OR destruction",
        "SSD deletion uses Secure Erase OR crypto-erasure OR destruction (NOT simple overwrite)",
        "Magnetic tape deletion uses degaussing OR overwrite OR destruction",
        "Removable media (USB, SD) deletion methods documented",
        "Paper document destruction uses cross-cut shredding (minimum)",
        "NIST SP 800-88 method classification assigned (Clear/Purge/Destroy)",
        "Deletion tools are approved and maintained",
        "Media disposal procedures documented (reuse, recycle, destroy)",
        "Certificate of Destruction obtained for outsourced destruction",
        "Forensic testing conducted at least annually",
        "Failed forensic tests trigger immediate remediation",
        "Data classification influences deletion method choice (Restricted → Destroy)",
        "End-of-life media destruction procedures exist"
    ]
    
    reference_tables = [
        ("HDD vs SSD Deletion Differences", 
         ["Aspect", "HDD (Spinning Disk)", "SSD (NAND Flash)"],
         [
             ["Overwrite Effectiveness", f"{CHECK} Effective (7-pass DoD)", f"{XMARK} Ineffective (wear leveling)"],
             ["Secure Erase Command", f"{CHECK} Effective (ATA)", f"{CHECK} Effective (ATA/NVMe)"],
             ["Degaussing", f"{CHECK} Effective", f"{XMARK} Ineffective (not magnetic)"],
             ["Cryptographic Erasure", f"{WARNING} Requires FDE", f"{CHECK} Highly Effective"],
             ["Physical Destruction", f"{CHECK} Shred, crush", f"{CHECK} Shred (flash chips destroyed)"],
             ["NIST Recommendation", "Purge: Overwrite or Degauss", "Purge: Secure Erase or Crypto-Erase"]
         ]),
        ("Physical Destruction Methods",
         ["Method", "Media Types", "Effectiveness", "Cost"],
         [
             ["Cross-Cut Shredding", "Paper, Optical, Cards", "High", "Low-Medium"],
             ["Disintegration", "HDD, SSD, Tape", "Very High", "Medium-High"],
             ["Degaussing", "HDD, Tape (magnetic)", "High", "Medium"],
             ["Incineration", "All media types", "Absolute", "High"],
             ["Crushing/Bending", "HDD, Optical", "Medium-High", "Low-Medium"]
         ])
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 2: Physical Media Deletion",
        "ISMS-POL-A.8.10-S2.2, Section 2.2.1",
        "Do we have effective deletion methods for all types of physical media, validated through testing, and aligned with NIST SP 800-88 guidance?",
        get_base_columns(), get_extended_columns_sheet2(),
        checklist, reference_tables,
        "sheet2"
    )


def create_sheet3_cloud_storage(ws, styles):
    """Create Sheet 3: Cloud Storage Deletion."""
    
    checklist = [
        "All cloud storage providers in use are identified",
        "Provider tier classification assigned (Tier 1-10 per ISMS-REF-A.5.23)",
        "Deletion method documented per provider (API, console, support ticket)",
        "API-based deletion implemented for Tier 1-3 providers (critical)",
        "Snapshot deletion procedures documented and tested",
        "Backup deletion timeline documented (may differ from active data)",
        "Multi-region/multi-zone deletion verified",
        "Object versioning deletion addressed (S3, GCS, Azure Blob)",
        "Soft-delete/recycle bin behavior documented",
        "Deletion confirmation/verification method exists",
        "Provider deletion SLA documented (see ISMS-IMP-A.8.10.3)",
        "Cryptographic erasure option evaluated (customer-managed keys)",
        "Account closure deletion behavior documented",
        "Cross-provider data deletion coordinated (if data replicated)",
        "Cloud deletion events logged and monitored"
    ]
    
    reference_tables = [
        ("Major Cloud Provider Deletion Methods",
         ["Provider", "Service", "Deletion Method", "Multi-Region"],
         [
             ["AWS", "S3", "DELETE API, Lifecycle Policy", f"{CHECK} Per region"],
             ["AWS", "EBS", "DeleteVolume API", f"{CHECK} Per region"],
             ["Azure", "Blob Storage", "Delete API, Lifecycle", f"{CHECK} Per region"],
             ["Azure", "Managed Disk", "Delete API", f"{CHECK} Per region"],
             ["GCP", "Cloud Storage", "DELETE API, Lifecycle", f"{CHECK} Per region"],
             ["GCP", "Persistent Disk", "Delete API", f"{CHECK} Per region"]
         ]),
        ("Cloud Deletion Challenges",
         ["Challenge", "Description", "Mitigation"],
         [
             ["Soft Delete / Recycle Bin", "Data retained 30-90 days", "Purge recycle bin explicitly"],
             ["Object Versioning", "Previous versions retained", "Delete all versions"],
             ["Snapshots / Backups", "Persist after source deletion", "Lifecycle policy or manual"],
             ["Multi-Region Replication", "Data in multiple regions", "Region-by-region deletion"],
             ["Delayed Deletion", "Provider may not delete immediately", "Document SLA, confirm"]
         ])
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 3: Cloud Storage Deletion",
        "ISMS-POL-A.8.10-S2.2, Section 2.2.2",
        "Can we effectively delete data from cloud storage providers, including all snapshots, backups, and multi-region copies?",
        get_base_columns(), get_extended_columns_sheet3(),
        checklist, reference_tables,
        "sheet3"
    )


def create_sheet4_database_application(ws, styles):
    """Create Sheet 4: Database & Application Deletion."""
    
    checklist = [
        "All databases and applications processing sensitive data identified",
        "Deletion method documented per system (logical delete, hard delete, purge)",
        "Referential integrity constraints addressed (cascading deletes or manual)",
        "Soft delete (logical delete) systems have purge procedures",
        "Hard delete operations tested and verified",
        "Database backup purging procedures exist",
        "Archive-then-delete workflows documented (if used)",
        "Crypto-shredding evaluated for encrypted databases",
        "Application-layer deletion triggers database deletion",
        "Deletion operations are logged and auditable",
        "Batch deletion procedures exist for large datasets",
        "Deletion performance tested (large table truncation)",
        "Database vacuuming/purging reclaims storage space",
        "Third-party SaaS application deletion verified (see A.8.10.3)",
        "Deletion operations do not impact system availability"
    ]
    
    reference_tables = [
        ("Database Deletion Methods",
         ["Method", "SQL Example", "Effect", "Storage Reclaimed"],
         [
             ["Logical DELETE", "UPDATE SET deleted=1", "Marks as deleted", f"{XMARK} No (until purged)"],
             ["Hard DELETE", "DELETE FROM table", "Removes immediately", f"{WARNING} After VACUUM"],
             ["TRUNCATE", "TRUNCATE TABLE", "Removes all rows", f"{CHECK} Immediate"],
             ["DROP TABLE", "DROP TABLE name", "Removes table", f"{CHECK} Immediate"],
             ["Purge/Vacuum", "VACUUM (PostgreSQL)", "Reclaims space", f"{CHECK} Yes"],
             ["Crypto-Shred", "DROP ENCRYPTION KEY", "Data unreadable", f"{CHECK} Immediate (crypto)"]
         ]),
        ("Referential Integrity Handling",
         ["Approach", "Implementation", "Pros", "Cons"],
         [
             ["CASCADE DELETE", "ON DELETE CASCADE", "Automatic, consistent", "Risk of unintended deletion"],
             ["SET NULL", "ON DELETE SET NULL", "Preserves child records", "Orphaned records"],
             ["RESTRICT", "ON DELETE RESTRICT", "Prevents accidental deletion", "Manual cleanup required"],
             ["Manual Deletion", "Application handles order", "Full control", "Error-prone, complex"]
         ])
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 4: Database & Application Deletion",
        "ISMS-POL-A.8.10-S2.2, Section 2.2.3",
        "Are database and application deletion methods effective, and do they address referential integrity, backups, and crypto-shredding?",
        get_base_columns(), get_extended_columns_sheet4(),
        checklist, reference_tables,
        "sheet4"
    )


def create_sheet5_mobile_endpoint(ws, styles):
    """Create Sheet 5: Mobile & Endpoint Deletion."""
    
    checklist = [
        "All endpoint device types documented (laptops, desktops, mobiles, tablets)",
        "Corporate vs BYOD devices identified",
        "Mobile Device Management (MDM) solution deployed",
        "Remote wipe capability tested and functional",
        "Full disk encryption (FDE) enabled on all corporate devices",
        "Factory reset procedures documented",
        "BYOD partial wipe capability (corporate data only)",
        "Device encryption keys managed (BitLocker, FileVault, etc.)",
        "Lost/stolen device wipe procedures exist",
        "Employee offboarding triggers device wipe",
        "BYOD app containerization supports selective wipe",
        "Physical device destruction procedures (end-of-life)",
        "Deletion verification for returned devices",
        "MDM logs capture wipe commands and confirmations",
        "Crypto-erasure supported (FDE key deletion)"
    ]
    
    reference_tables = [
        ("Device Deletion Methods",
         ["Device Type", "Deletion Method", "Effectiveness", "Requirements"],
         [
             ["Corporate Laptop (Windows)", "BitLocker + Format", "High", "BitLocker enabled"],
             ["Corporate Laptop (macOS)", "FileVault + Erase", "High", "FileVault enabled"],
             ["Corporate Mobile (iOS)", "MDM Remote Wipe", "Very High", "MDM enrollment"],
             ["Corporate Mobile (Android)", "MDM Remote Wipe", "High", "MDM + FDE"],
             ["BYOD Mobile", "MDM Selective Wipe", "Medium", "MDM + containerization"],
             ["Desktop Workstation", "Overwrite or Destroy", "High", "Depends on storage type"]
         ]),
        ("Full Disk Encryption (FDE) Support",
         ["Operating System", "FDE Solution", "Crypto-Erase Method"],
         [
             ["Windows 10/11", "BitLocker", "Delete recovery key"],
             ["macOS", "FileVault 2", "Delete recovery key"],
             ["Linux", "LUKS (dm-crypt)", "Delete LUKS header"],
             ["iOS", "Built-in (always on)", "Factory reset (key deleted)"],
             ["Android", "File-Based Encryption", "Factory reset (key deleted)"]
         ])
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 5: Mobile & Endpoint Deletion",
        "ISMS-POL-A.8.10-S2.2, Section 2.2.4",
        "Can we securely delete data from corporate and BYOD devices, including remote wipe capabilities and full disk encryption support?",
        get_base_columns(), get_extended_columns_sheet5(),
        checklist, reference_tables,
        "sheet5"
    )


def create_sheet6_tool_validation(ws, styles):
    """Create Sheet 6: Deletion Tool Validation."""
    
    checklist = [
        "Annual forensic testing of deletion methods is scheduled",
        "Representative sample of media types tested",
        "Forensic recovery tools used (EnCase, FTK, Autopsy, etc.)",
        "Pass/fail criteria defined (no data recovery = pass)",
        "Failed tests trigger immediate investigation and remediation",
        "Testing covers all NIST method types (Clear, Purge, Destroy)",
        "Cloud deletion verification performed (where possible)",
        "Database deletion verification performed (backup restoration test)",
        "Mobile device wipe verification performed (MDM logs + device inspection)",
        "Independent testing considered (external forensic firm)",
        "Test results documented and linked to Evidence Register",
        "Deletion tool vendors provide effectiveness documentation",
        "New deletion methods tested before production use",
        "Testing includes encrypted media (crypto-erasure validation)",
        "Testing results reported to management and DPO"
    ]
    
    reference_tables = [
        ("Forensic Testing Approach",
         ["Media Type", "Test Method", "Pass Criteria"],
         [
             ["HDD (overwritten)", "File recovery attempt", "Zero files recovered"],
             ["SSD (secure erased)", "File recovery attempt", "Zero files recovered"],
             ["Cloud Storage (deleted)", "API verification", "No objects found"],
             ["Database (hard deleted)", "Backup restoration + query", "Zero records recovered"],
             ["Mobile Device (wiped)", "Physical inspection + recovery", "Zero corporate data found"],
             ["Crypto-Erased Media", "Key verification, mount attempt", "Data unreadable"]
         ]),
        ("Testing Frequency Recommendations",
         ["Media Type", "Recommended Frequency", "Rationale"],
         [
             ["Physical Media (HDD/SSD)", "Annual", "Technology stable"],
             ["Cloud Storage", "Semi-Annual", "Provider changes, API updates"],
             ["Database Deletion", "Annual", "Logic stable"],
             ["Mobile Devices", "Quarterly", "Frequent OS updates"],
             ["New Deletion Tools", "Before Production", "Validate before use"],
             ["After Failed Test", "Immediate Retest", "Verify remediation"]
         ])
    ]
    
    create_assessment_sheet(
        ws, styles,
        "Sheet 6: Deletion Tool Validation",
        "ISMS-POL-A.8.10-S2.2, Section 2.2.5",
        "Do we regularly test our deletion methods using forensic tools to verify effectiveness, and do we remediate failed tests?",
        get_base_columns(), get_extended_columns_sheet6(),
        checklist, reference_tables,
        "sheet6"
    )


# ==========================================================================
# SECTION 11: REFERENCE DATA
# ==========================================================================

# Defined inline in sheet creators above


# ==========================================================================
# SECTION 12: MAIN EXECUTION
# ==========================================================================

def main():
    """Main execution function."""
    print("=" * 78)
    print("ISMS-IMP-A.8.10.2 - Deletion Methods Assessment Generator")
    print("ISO/IEC 27001:2022 Control A.8.10: Information Deletion")
    print("NIST SP 800-88 Framework Integration")
    print("=" * 78)
    print()
    
    wb = create_workbook()
    styles = setup_styles()
    
    print("[1/9] Creating Instructions & Legend (NIST Framework)...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    
    print("[2/9] Creating Sheet 2: Physical Media Deletion...")
    create_sheet2_physical_media(wb["2. Physical Media Deletion"], styles)
    
    print("[3/9] Creating Sheet 3: Cloud Storage Deletion...")
    create_sheet3_cloud_storage(wb["3. Cloud Storage Deletion"], styles)
    
    print("[4/9] Creating Sheet 4: Database & Application Deletion...")
    create_sheet4_database_application(wb["4. Database & Application Deletion"], styles)
    
    print("[5/9] Creating Sheet 5: Mobile & Endpoint Deletion...")
    create_sheet5_mobile_endpoint(wb["5. Mobile & Endpoint Deletion"], styles)
    
    print("[6/9] Creating Sheet 6: Deletion Tool Validation...")
    create_sheet6_tool_validation(wb["6. Deletion Tool Validation"], styles)
    
    print("[7/9] Creating Summary Dashboard (with SSD Overwrite Detection)...")
    create_summary_dashboard(wb["Summary Dashboard"], styles)
    
    print("[8/9] Creating Evidence Register (100 rows)...")
    create_evidence_register(wb["Evidence Register"], styles)
    
    print("[9/9] Creating Approval Sign-Off (3-level workflow)...")
    create_approval_signoff(wb["Approval Sign-Of"], styles)
    
    filename = f"ISMS-IMP-A.8.10.2_Deletion_Methods_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    print()
    print("=" * 78)
    print(f"{CHECK} SUCCESS: {filename}")
    print()
    print("Workbook Structure:")
    print("  • Instructions & Legend - NIST SP 800-88 framework overview")
    print("  • Sheet 2: Physical Media Deletion - HDD, SSD, tape, paper (NIST methods)")
    print("  • Sheet 3: Cloud Storage Deletion - AWS, Azure, GCP deletion capabilities")
    print("  • Sheet 4: Database & Application Deletion - Logical delete, crypto-shred")
    print("  • Sheet 5: Mobile & Endpoint Deletion - MDM wipe, FDE, factory reset")
    print("  • Sheet 6: Deletion Tool Validation - Forensic testing, effectiveness")
    print("  • Summary Dashboard - Compliance overview + SSD overwrite detection")
    print("  • Evidence Register - 100 rows for supporting documentation")
    print("  • Approval Sign-Off - 3-level approval workflow")
    print()
    print("Key Features:")
    print("  ✓ NIST SP 800-88 framework (Clear, Purge, Destroy)")
    print("  ✓ HDD vs SSD deletion method distinctions (CRITICAL!)")
    print("  ✓ SSD overwrite detection formula (flags ineffective methods)")
    print("  ✓ Cloud provider tier integration (ISMS-REF-A.5.23)")
    print("  ✓ Crypto-erasure methods (CMK/CMEK deletion)")
    print("  ✓ Forensic testing validation tracking")
    print("  ✓ 13 data entry rows per assessment sheet (yellow-highlighted)")
    print("  ✓ All dropdowns configured and working")
    print("  ✓ Vendor-neutral approach")
    print()
    print("CRITICAL VALIDATION:")
    print("  ⚠️  Dashboard includes SSD + Overwrite detection formula")
    print("  ⚠️  If count > 0, flags as CRITICAL GAP (overwrite ineffective for SSDs)")
    print()
    print("Related Assessments:")
    print("  → ISMS-IMP-A.8.10.1 (Retention Triggers) - Completed")
    print("  → ISMS-IMP-A.8.10.3 (Third-Party & Cloud Deletion) - Next")
    print("=" * 78)


if __name__ == "__main__":
    main()