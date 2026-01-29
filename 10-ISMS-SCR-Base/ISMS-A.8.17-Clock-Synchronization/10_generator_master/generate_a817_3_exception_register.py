#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.17.3 - Time Synchronization Exception Register Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.17: Clock Synchronization
Exception Management: Non-Compliant Systems and Risk Acceptance

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific exception management processes, risk acceptance
workflows, and compliance tracking requirements.

Key customization areas:
1. Exception approval workflow (match your governance structure)
2. Risk scoring criteria (adapt to your risk assessment methodology)
3. Compensating controls (specific to your security architecture)
4. Review cycles and escalation procedures (based on your policies)
5. Integration with GRC systems (if applicable)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.17 Clock Synchronization Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel exception register for documenting
and managing systems that cannot meet A.8.17 time synchronization requirements,
with formal risk acceptance and compensating controls.

**Purpose:**
Enables systematic tracking of exceptions to time synchronization requirements,
including documented business justification, risk assessment, compensating
controls, and time-bound risk acceptance with defined review cycles.

**Exception Scenarios:**
- Air-gapped systems without network connectivity
- Legacy systems incompatible with modern NTP protocols
- IoT/embedded devices with limited time sync capabilities
- Systems with operational constraints preventing time sync
- Temporary exceptions during infrastructure migrations
- Systems undergoing remediation with interim risk acceptance

**Generated Workbook Structure:**
1. Instructions & Legend - Exception process guidance and criteria
2. Exception_Register - Active exception tracking and status
3. Risk_Assessment - Risk scoring and impact analysis per exception
4. Compensating_Controls - Alternative controls for each exception
5. Review_Schedule - Exception review and re-approval tracking
6. Expired_Exceptions - Overdue exceptions requiring action
7. Remediation_Plan - Path to compliance for temporary exceptions
8. Evidence_Register - Supporting documentation linkage
9. Approval & Sign-Off - Risk owner and CISO approval workflow

**Key Features:**
- Structured exception documentation with business justification
- Risk scoring based on system criticality and exposure
- Compensating control identification and validation
- Time-bound exceptions with automatic expiry tracking
- Review cycle enforcement
- Approval workflow with multi-level sign-off
- Exception statistics and trending
- Integration with gap analysis from Assessment 2

**Integration:**
Exception register data feeds into the A.8.17 Compliance Dashboard to provide
complete compliance picture including both compliant systems and formally
accepted exceptions.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

Optional:
    - Integration with GRC platforms for exception workflow
    - Import gaps from Assessment 2 (System Synchronization Status)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a817_3_exception_register.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a817_3_exception_register.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a817_3_exception_register.py --date 20250125
    
    # Import gaps from Assessment 2
    python3 generate_a817_3_exception_register.py --import-gaps ISMS-A.8.17-Assessment-2.xlsx

Output:
    File: ISMS-A.8.17-Assessment-3-Exception-Register-YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Import or manually enter systems requiring exceptions
    2. Document business justification for each exception
    3. Conduct risk assessment (likelihood × impact)
    4. Identify and validate compensating controls
    5. Define exception duration and review cycle
    6. Obtain risk owner approval
    7. Obtain CISO/security approval
    8. Schedule exception review meetings
    9. Monitor for expired exceptions
    10. Feed results into A.8.17 Compliance Dashboard

Exception Approval Process:
    1. System owner requests exception with justification
    2. Security team assesses risk and proposes compensating controls
    3. Risk owner reviews and approves/rejects
    4. CISO provides final approval for high-risk exceptions
    5. Exception is time-bound with defined review cycle
    6. Reviews ensure exception remains valid or is remediated

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.17
Assessment Domain:    3 of 2 (Exception Management - Optional)
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.17: Clock Synchronization Policy (Requirements)
    - ISMS-IMP-A.8.17.2: System Synchronization Status (Gap Source)
    - A.8.17 Compliance Dashboard (Consolidation)
    - Corporate Risk Management Policy (Exception Approval Process)

Related Standards:
    - ISO/IEC 27001:2022 Clause 6.1.3: Risk Treatment
    - ISO/IEC 27002:2022: Risk Acceptance Guidance

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements exception tracking and risk acceptance workflow
    - Supports compensating control documentation
    - Time-bound exceptions with automatic expiry tracking
    - Integrated with A.8.17 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Exception Philosophy:**
Exceptions are not "free passes" - they represent formally accepted risk with:
- Clear business justification
- Documented risk assessment
- Defined compensating controls
- Time-bound approval
- Regular review cycles
- Path to remediation where feasible

**Risk Assessment:**
Each exception must include risk scoring based on:
- System criticality (impact if compromised)
- Data sensitivity handled by system
- Exposure to security events
- Likelihood of time-related security issues
- Existing compensating controls

Risk scores inform approval levels and review frequency.

**Compensating Controls:**
Acceptable compensating controls may include:
- Enhanced logging/monitoring for affected system
- Manual log correlation procedures
- Physical security controls (air-gapped systems)
- Network segmentation isolating affected system
- Reduced system criticality/scope

Controls must demonstrably reduce risk to acceptable level.

**Review Cycles:**
Exception review frequency based on risk:
- High risk: Quarterly review required
- Medium risk: Semi-annual review
- Low risk: Annual review
- All exceptions: Must be renewed annually at minimum

**Audit Considerations:**
Auditors will scrutinize exceptions for:
- Adequate business justification
- Appropriate risk assessment
- Effective compensating controls
- Current approvals (not expired)
- Evidence of regular reviews
- Remediation progress for temporary exceptions

Ensure exception register is audit-ready with complete documentation.

**Integration with Assessment 2:**
Gap Analysis sheet in Assessment 2 (System Sync Status) identifies systems
not meeting requirements. These gaps feed into exception register where:
- Exception is formally requested
- Risk is assessed and accepted
- OR remediation is planned and executed

**Expired Exceptions:**
Systems with expired exceptions must be either:
- Remediated to compliance
- Re-evaluated and exception renewed
- Decommissioned if no longer needed

Never allow exceptions to remain active past expiry - this is audit finding.

**Approval Authority:**
Define clear approval levels:
- System owner: Initial exception request
- Risk owner: Risk acceptance for business unit
- Security team: Compensating control validation
- CISO: High-risk exceptions (>70 risk score)
- Executive sponsor: Exceptions >90 days duration

**Data Protection:**
Exception register contains sensitive information including:
- System security weaknesses
- Risk assessment details
- Infrastructure vulnerabilities
- Compensating control strategies

Handle as CONFIDENTIAL per organizational data classification.

**Quality Assurance:**
Security governance team should audit exception register quarterly for:
- Expired exceptions requiring action
- Exceptions without adequate compensating controls
- Patterns indicating systemic issues
- Opportunities for infrastructure improvement

================================================================================
"""

import argparse
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation



# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CLOCK = '\u23F0'      # ⏰ Alarm clock
SYNC = '\U0001F504'   # 🔄 Counterclockwise arrows
HOURGLASS = '\u23F3'  # ⏳ Hourglass
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
BULLET = '\u2022'     # • Bullet
ARROW = '\u2192'      # → Right arrow

def create_styles():
    """Define A.8.24 standard styles for the workbook"""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "data": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "title": {
            "font": Font(name="Calibri", bold=True, size=16, color="003366"),
            "alignment": Alignment(horizontal="left", vertical="center")
        },
        "subtitle": {
            "font": Font(name="Calibri", bold=True, size=12, color="003366"),
            "alignment": Alignment(horizontal="left", vertical="center")
        },
        "warning": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "good": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }

def set_column_widths(ws, widths):
    """Set column widths for a worksheet"""
    for col_num, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

def apply_style(cell, style_dict):
    """Apply style dictionary to cell"""
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

def create_instructions_sheet(wb, styles):
    """Create instructions sheet"""
    ws = wb["Instructions"]
    
    # Title
    ws['A1'] = "ISMS A.8.17 - Clock Synchronization Exception Register"
    apply_style(ws['A1'], styles['title'])
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:H2')
    
    # Document ID
    ws['A4'] = "Document ID:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "ISMS-A.8.17-Exception-Register"
    ws['B4'].font = Font(bold=True, color="003366")
    
    ws['A5'] = "Policy Reference:"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = "ISMS-POL-A.8.17 Section 3.3 (Exception Management)"
    
    # Instructions
    row = 7
    instructions = [
        ("Purpose", "Track and manage exceptions to clock synchronization policy requirements"),
        ("", ""),
        ("What Requires an Exception?", ""),
        ("  • Systems unable to synchronize (air-gapped, vendor limitations)", ""),
        ("  • Alternative time sources outside policy requirements", ""),
        ("  • Temporary non-compliance during remediation", ""),
        ("  • Policy requirement waivers", ""),
        ("", ""),
        ("Approval Authority (per Policy Section 3.3):", ""),
        ("  • Technical Exceptions → CISO approval", ""),
        ("  • Policy-Level Exceptions → Executive Management approval", ""),
        ("  • Maximum Duration: 12 months for temporary exceptions", ""),
        ("", ""),
        ("Required Information:", ""),
        ("  1. Exception ID (auto-generated: EXC-A817-001, EXC-A817-002, etc.)", ""),
        ("  2. System/Asset Name requiring exception", ""),
        ("  3. Specific requirement exempted (REQ-817-xxx from policy)", ""),
        ("  4. Business justification", ""),
        ("  5. Risk assessment (impact if control fails)", ""),
        ("  6. Compensating controls (how risk is mitigated)", ""),
        ("  7. Approval authority and date", ""),
        ("  8. Expiry date (max 12 months)", ""),
        ("", ""),
        ("Quarterly Review Process:", ""),
        ("  • Review active exceptions every quarter", ""),
        ("  • Verify compensating controls still effective", ""),
        ("  • Check if compliance has become feasible", ""),
        ("  • Revoke if risk profile changed", ""),
        ("", ""),
        ("90-Day Reassessment Rule:", ""),
        ("  • Exceptions granted within 90 days of annual policy review", ""),
        ("    must be reassessed during that review cycle", ""),
        ("  • Reassess against CURRENT requirements, not original approval", ""),
        ("", ""),
        ("Workflow:", ""),
        ("  1. Complete Exception_Requests sheet (one row per exception)", ""),
        ("  2. CISO or Executive Management reviews and approves", ""),
        ("  3. Move to Active_Exceptions sheet", ""),
        ("  4. Quarterly: Review active exceptions, update status", ""),
        ("  5. When expired/revoked: Move to Expired_Exceptions sheet", ""),
    ]
    
    for label, value in instructions:
        ws[f'A{row}'] = label
        if value:
            ws[f'B{row}'] = value
        if label and not value:
            ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    set_column_widths(ws, [35, 80])

def create_exception_requests_sheet(wb, styles):
    """Create exception requests sheet (pending approval)"""
    ws = wb["Exception_Requests"]
    
    # Title
    ws['A1'] = "Exception Requests (Pending Approval)"
    apply_style(ws['A1'], styles['subtitle'])
    ws.merge_cells('A1:M1')
    
    # Headers
    headers = [
        "Exception ID [*]",
        "Submitted Date [*]",
        "System/Asset Name [*]",
        "Requirement Exempted [*]",
        "Exception Type [*]",
        "Business Justification [*]",
        "Risk Assessment [*]",
        "Compensating Controls [*]",
        "Requested Duration [*]",
        "Submitted By [*]",
        "Approval Authority [*]",
        "Approval Status",
        "Notes",
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        apply_style(cell, styles['header'])
    
    # Data validations
    exception_type_dv = DataValidation(
        type="list",
        formula1='"Technical Exception,Policy-Level Exception,Temporary Non-Compliance"',
        allow_blank=False
    )
    ws.add_data_validation(exception_type_dv)
    
    approval_auth_dv = DataValidation(
        type="list",
        formula1='"CISO,Executive Management"',
        allow_blank=False
    )
    ws.add_data_validation(approval_auth_dv)
    
    approval_status_dv = DataValidation(
        type="list",
        formula1='"Pending,Approved,Rejected,Withdrawn"',
        allow_blank=False
    )
    ws.add_data_validation(approval_status_dv)
    
    # Example row
    example = [
        "EXC-A817-001",
        datetime.now().strftime('%Y-%m-%d'),
        "legacy-scada-system.example.com",
        "REQ-817-009 (All systems must synchronize)",
        "Technical Exception",
        "Air-gapped SCADA system cannot reach external NTP servers; vendor does not support NTP protocol",
        "Moderate risk: Time drift may cause log correlation issues; no authentication impact (isolated network)",
        "GPS time receiver installed for Stratum 0 local time source; manual time verification weekly",
        "12 months (temporary until vendor upgrade)",
        "John Smith (SCADA Admin)",
        "CISO",
        "Pending",
        "Vendor upgrade planned Q3 2026",
    ]
    
    for col_num, value in enumerate(example, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = value
        apply_style(cell, styles['data'])
        cell.fill = styles['input_cell']['fill']
    
    # Apply validations to data rows (4-53)
    for row in range(4, 54):
        exception_type_dv.add(ws.cell(row=row, column=5))
        approval_auth_dv.add(ws.cell(row=row, column=11))
        approval_status_dv.add(ws.cell(row=row, column=12))
        
        # Auto-generate Exception ID
        if row > 4:
            ws.cell(row=row, column=1).value = f"EXC-A817-{row-3:03d}"
            ws.cell(row=row, column=1).font = Font(color="808080")
        
        # Apply borders to all cells
        for col in range(1, 14):
            if ws.cell(row=row, column=col).value is None:
                cell = ws.cell(row=row, column=col)
                apply_style(cell, styles['data'])
                cell.fill = styles['input_cell']['fill']
    
    set_column_widths(ws, [18, 15, 30, 25, 22, 40, 40, 40, 20, 25, 20, 18, 30])
    ws.freeze_panes = 'A4'

def create_active_exceptions_sheet(wb, styles):
    """Create active exceptions sheet (approved and in effect)"""
    ws = wb["Active_Exceptions"]
    
    # Title
    ws['A1'] = "Active Exceptions (Approved)"
    apply_style(ws['A1'], styles['subtitle'])
    ws.merge_cells('A1:O1')
    
    # Headers
    headers = [
        "Exception ID [*]",
        "System/Asset Name [*]",
        "Requirement Exempted [*]",
        "Exception Type [*]",
        "Approved By [*]",
        "Approval Date [*]",
        "Expiry Date [*]",
        "Days Until Expiry",
        "Compensating Controls [*]",
        "Last Quarterly Review",
        "Next Review Due",
        "Reassessment Required",
        "Status",
        "Review Notes",
        "Actions Required",
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        apply_style(cell, styles['header'])
    
    # Data validations
    exception_type_dv = DataValidation(
        type="list",
        formula1='"Technical Exception,Policy-Level Exception,Temporary Non-Compliance"',
        allow_blank=False
    )
    ws.add_data_validation(exception_type_dv)
    
    reassess_dv = DataValidation(
        type="list",
        formula1='"Yes - Within 90 days of policy review,No"',
        allow_blank=False
    )
    ws.add_data_validation(reassess_dv)
    
    status_dv = DataValidation(
        type="list",
        formula1='"Active,Expiring Soon,Under Review,Pending Revocation"',
        allow_blank=False
    )
    ws.add_data_validation(status_dv)
    
    # Example row
    today = datetime.now()
    expiry = today + timedelta(days=180)
    next_review = today + timedelta(days=90)
    
    example = [
        "EXC-A817-001",
        "legacy-scada-system.example.com",
        "REQ-817-009 (All systems must synchronize)",
        "Technical Exception",
        "Jane Doe (CISO)",
        today.strftime('%Y-%m-%d'),
        expiry.strftime('%Y-%m-%d'),
        f"=G4-TODAY()",  # Days until expiry formula
        "GPS time receiver (Stratum 0); Weekly manual verification",
        today.strftime('%Y-%m-%d'),
        next_review.strftime('%Y-%m-%d'),
        "No",
        "Active",
        "Compensating controls verified; GPS receiver operational",
        "Continue quarterly reviews; Monitor vendor upgrade timeline",
    ]
    
    for col_num, value in enumerate(example, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = value
        apply_style(cell, styles['data'])
        if col_num == 8:  # Days until expiry
            cell.font = Font(color="0000FF")
        else:
            cell.fill = styles['input_cell']['fill']
    
    # Apply validations and formulas to data rows (4-53)
    for row in range(4, 54):
        exception_type_dv.add(ws.cell(row=row, column=4))
        reassess_dv.add(ws.cell(row=row, column=12))
        status_dv.add(ws.cell(row=row, column=13))
        
        # Days until expiry formula
        if row > 4:
            ws.cell(row=row, column=8).value = f"=IF(ISBLANK(G{row}),\"\",G{row}-TODAY())"
            ws.cell(row=row, column=8).font = Font(color="0000FF")
            
            # Next review due (90 days from last review)
            ws.cell(row=row, column=11).value = f"=IF(ISBLANK(J{row}),\"\",J{row}+90)"
            ws.cell(row=row, column=11).font = Font(color="0000FF")
        
        # Apply borders
        for col in range(1, 16):
            if ws.cell(row=row, column=col).value is None or col in [8, 11]:
                if col not in [8, 11]:  # Don't overwrite formulas
                    cell = ws.cell(row=row, column=col)
                    apply_style(cell, styles['data'])
                    cell.fill = styles['input_cell']['fill']
    
    # Conditional formatting note
    ws['A56'] = "NOTE: Exceptions expiring within 30 days require renewal or revocation decision"
    ws['A56'].font = Font(italic=True, color="FF0000")
    ws.merge_cells('A56:O56')
    
    set_column_widths(ws, [18, 30, 30, 22, 25, 15, 15, 18, 40, 18, 18, 22, 18, 40, 40])
    ws.freeze_panes = 'A4'

def create_expired_exceptions_sheet(wb, styles):
    """Create expired/revoked exceptions sheet (historical record)"""
    ws = wb["Expired_Exceptions"]
    
    # Title
    ws['A1'] = "Expired and Revoked Exceptions (Historical Record)"
    apply_style(ws['A1'], styles['subtitle'])
    ws.merge_cells('A1:L1')
    
    # Headers
    headers = [
        "Exception ID",
        "System/Asset Name",
        "Requirement Exempted",
        "Exception Type",
        "Approved By",
        "Approval Date",
        "Expiry Date",
        "Closure Date",
        "Closure Reason",
        "Total Duration (Days)",
        "Final Status",
        "Lessons Learned",
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        apply_style(cell, styles['header'])
    
    # Data validations
    closure_reason_dv = DataValidation(
        type="list",
        formula1='"Expired - Not Renewed,Compliance Achieved,Revoked - Risk Changed,Revoked - Controls Failed,System Decommissioned"',
        allow_blank=False
    )
    ws.add_data_validation(closure_reason_dv)
    
    final_status_dv = DataValidation(
        type="list",
        formula1='"Completed Successfully,Revoked,Replaced,System Retired"',
        allow_blank=False
    )
    ws.add_data_validation(final_status_dv)
    
    # Example row
    example = [
        "EXC-A817-002",
        "old-db-server.example.com",
        "REQ-817-011 (Drift threshold compliance)",
        "Temporary Non-Compliance",
        "Jane Doe (CISO)",
        "2025-06-15",
        "2025-12-15",
        "2025-11-30",
        "Compliance Achieved",
        "=H4-F4",  # Duration formula
        "Completed Successfully",
        "NTP configuration corrected; System now compliant with drift thresholds",
    ]
    
    for col_num, value in enumerate(example, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = value
        apply_style(cell, styles['data'])
        if col_num == 10:  # Duration formula
            cell.font = Font(color="0000FF")
        else:
            cell.fill = styles['input_cell']['fill']
    
    # Apply validations to data rows (4-103)
    for row in range(4, 104):
        closure_reason_dv.add(ws.cell(row=row, column=9))
        final_status_dv.add(ws.cell(row=row, column=11))
        
        # Duration formula
        if row > 4:
            ws.cell(row=row, column=10).value = f"=IF(OR(ISBLANK(F{row}),ISBLANK(H{row})),\"\",H{row}-F{row})"
            ws.cell(row=row, column=10).font = Font(color="0000FF")
        
        # Apply borders
        for col in range(1, 13):
            if ws.cell(row=row, column=col).value is None or col == 10:
                if col != 10:  # Don't overwrite formula
                    cell = ws.cell(row=row, column=col)
                    apply_style(cell, styles['data'])
                    cell.fill = styles['input_cell']['fill']
    
    set_column_widths(ws, [18, 30, 30, 22, 25, 15, 15, 15, 25, 20, 20, 50])
    ws.freeze_panes = 'A4'

def create_summary_dashboard_sheet(wb, styles):
    """Create exception summary dashboard"""
    ws = wb["Summary_Dashboard"]
    
    # Title
    ws['A1'] = "Exception Register Summary Dashboard"
    apply_style(ws['A1'], styles['title'])
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(italic=True)
    ws.merge_cells('A2:F2')
    
    # Key Metrics
    row = 4
    ws[f'A{row}'] = "Exception Summary"
    apply_style(ws[f'A{row}'], styles['subtitle'])
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    metrics = [
        ("Pending Approval", "=COUNTA(Exception_Requests!A4:A53)"),
        ("Active Exceptions", "=COUNTA(Active_Exceptions!A4:A53)"),
        ("Expired/Revoked (Total)", "=COUNTA(Expired_Exceptions!A4:A103)"),
        ("Expiring Within 30 Days", "=COUNTIF(Active_Exceptions!H4:H53,\"<30\")"),
        ("Requiring Reassessment (90-day rule)", "=COUNTIF(Active_Exceptions!L4:L53,\"Yes*\")"),
    ]
    
    for metric_name, formula in metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = formula
        ws[f'B{row}'].font = Font(bold=True, size=14, color="003366")
        ws[f'B{row}'].alignment = Alignment(horizontal="center")
        row += 1
    
    # Alerts
    row += 2
    ws[f'A{row}'] = "Alerts & Action Items"
    apply_style(ws[f'A{row}'], styles['subtitle'])
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    alerts = [
        "Exceptions expiring within 30 days require renewal decision",
        "Quarterly reviews overdue if 'Next Review Due' date has passed",
        "Exceptions marked for reassessment must be reviewed in next policy cycle",
        "All pending approvals should be processed within 30 days of submission",
    ]
    
    for alert in alerts:
        ws[f'A{row}'] = f"{BULLET} {alert}"
        row += 1
    
    set_column_widths(ws, [40, 20, 20, 20, 20, 20])

def main():
    """Generate exception register workbook"""
    parser = argparse.ArgumentParser(
        description='Generate ISMS A.8.17 Exception Register workbook'
    )
    parser.add_argument(
        '--output',
        default=f'ISMS-IMP-A.8.17.3_Exception_Register_{datetime.now().strftime("%Y%m%d")}.xlsx',
        help='Output filename (default: ISMS-A.8.17-Exception-Register_YYYYMMDD.xlsx)'
    )
    args = parser.parse_args()
    
    print("="*80)
    print("ISMS A.8.17 - Clock Synchronization Exception Register Generator")
    print("="*80)
    print()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create styles
    styles = create_styles()
    
    # Create sheets
    print("Creating sheets...")
    wb.create_sheet("Instructions")
    wb.create_sheet("Exception_Requests")
    wb.create_sheet("Active_Exceptions")
    wb.create_sheet("Expired_Exceptions")
    wb.create_sheet("Summary_Dashboard")
    
    print("  [1/5] Instructions...")
    create_instructions_sheet(wb, styles)
    
    print("  [2/5] Exception Requests...")
    create_exception_requests_sheet(wb, styles)
    
    print("  [3/5] Active Exceptions...")
    create_active_exceptions_sheet(wb, styles)
    
    print("  [4/5] Expired Exceptions...")
    create_expired_exceptions_sheet(wb, styles)
    
    print("  [5/5] Summary Dashboard...")
    create_summary_dashboard_sheet(wb, styles)
    
    # Save workbook
    wb.save(args.output)
    
    print()
    print("="*80)
    print(f"{CHECK} SUCCESS: {args.output}")
    print("="*80)
    print()
    print("Exception Register Structure:")
    print("  • Instructions - Usage guidance and policy references")
    print("  • Exception_Requests - Pending approvals")
    print("  • Active_Exceptions - Approved exceptions in effect")
    print("  • Expired_Exceptions - Historical record")
    print("  • Summary_Dashboard - Metrics and alerts")
    print()
    print("Next Steps:")
    print("  1. Open workbook and review Instructions sheet")
    print("  2. Complete Exception_Requests for new exceptions")
    print("  3. CISO/Executive reviews and approves")
    print("  4. Move approved to Active_Exceptions")
    print("  5. Quarterly review of all active exceptions")
    print("  6. Move expired/revoked to Expired_Exceptions")
    print()
    print("Policy Reference: ISMS-POL-A.8.17 Section 3.3 (Exception Management)")
    print()

if __name__ == "__main__":
    main()
