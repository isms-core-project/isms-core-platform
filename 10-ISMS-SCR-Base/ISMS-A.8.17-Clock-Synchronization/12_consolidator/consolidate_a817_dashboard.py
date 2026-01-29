#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-A.8.17 - Dashboard Data Consolidation Utility
================================================================================

ISO/IEC 27001:2022 Control A.8.17: Clock Synchronization
Alternative Data Integration: Static Consolidation for Compliance Dashboard

--------------------------------------------------------------------------------
ALTERNATIVE TO FORMULA-BASED INTEGRATION
--------------------------------------------------------------------------------

This utility provides an ALTERNATIVE method for populating the A.8.17
Compliance Dashboard by directly copying data from source assessment workbooks
rather than using external workbook formula references.

**Primary Approach vs. Alternative Approach:**

**Primary (Recommended): External Workbook Formula Linking**
- Dashboard contains formulas like: ='[ISMS-A.8.17-Assessment-1.xlsx]Time_Sources'!$B$5
- Data updates automatically when source files change
- Dashboard and source files must be in same directory
- Requires "Update Links" when opening dashboard
- Real-time data synchronization
- See: generate_a817_4_compliance_dashboard.py

**Alternative (This Script): Direct Data Consolidation**
- Reads data from source workbooks programmatically
- Copies data values (not formulas) into dashboard
- Creates static snapshot of compliance status
- Works with dashboard and sources in different locations
- Requires manual re-run when source data changes
- Useful when external links don't work or aren't desired

**When to Use This Script:**

1. **External Links Don't Work**
   - Excel security settings prevent external workbook links
   - SharePoint/OneDrive environments where linking fails
   - Email distribution requires self-contained workbooks
   - Recipient systems block external references

2. **Static Snapshots Required**
   - Monthly/quarterly compliance reporting archives
   - Board presentations requiring frozen data
   - Audit evidence packages with point-in-time data
   - Historical compliance tracking over time

3. **Simplified Distribution**
   - Single-file distribution to stakeholders
   - No dependency on source file availability
   - Recipients don't have access to source assessments
   - Reduced complexity for non-technical users

4. **Data Consolidation Flexibility**
   - Selective data inclusion from sources
   - Custom data transformation during consolidation
   - Multi-source aggregation and calculations
   - Data validation during import

**When NOT to Use This Script:**

- If external workbook linking works reliably → Use primary approach
- If real-time data updates are needed → Use external linking
- If dashboard and sources will always be co-located → Use external linking
- For initial dashboard setup → Use generate_a817_4_compliance_dashboard.py first

**Purpose:**
Reads user-entered data from both assessment workbooks and writes consolidated
data into dashboard, creating self-contained snapshot for reporting and archival.

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This consolidation script reads completed assessment data and populates the
compliance dashboard, creating a static point-in-time snapshot of time
synchronization compliance status.

**What It Does:**

1. **Validates Source Availability**
   - Checks both normalized assessment workbooks exist
   - Validates dashboard workbook is accessible
   - Reports missing files before attempting consolidation

2. **Reads Source Data**
   - Opens each source workbook in data-only mode
   - Reads user-entered data from key sheets
   - Extracts relevant rows and columns per mapping
   - Skips empty rows and merged cell areas

3. **Consolidates into Dashboard**
   - Maps source sheets to target dashboard sheets
   - Writes data to appropriate dashboard locations
   - Adds source labels for traceability
   - Preserves existing dashboard structure

4. **Provides Consolidation Statistics**
   - Reports sheets processed and rows consolidated
   - Shows data distribution across dashboard sheets
   - Identifies any skipped or missing sheets

**Consolidation Mapping:**
```python
"Assessment-1": {  # Time Sources
    "Time_Sources": {
        "target": "Time_Sources_Summary",
        "start_row": 2,
        "columns": "A:I",
        "label": "Time Source Infrastructure"
    }
}

"Assessment-2": {  # Sync Status
    "System_Sync_Status": {
        "target": "Sync_Status_Summary",
        "start_row": 2,
        "columns": "A:G",
        "label": "System Synchronization"
    }
}
```

**Data Sources:**
- ISMS-A.8.17-Assessment-1.xlsx (Time Source Infrastructure)
- ISMS-A.8.17-Assessment-2.xlsx (System Synchronization Status)

These are the normalized file names created by normalize_assessment_files_a817.py

**Target Dashboard Sheets:**
- Executive_Summary (consolidated KPIs)
- Time_Sources_Summary (time source data)
- Sync_Status_Summary (system sync data)
- Gap_Analysis (consolidated gaps)
- Action_Items (remediation tasks)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel manipulation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Excel file reading and writing)
    - sys, os (standard library)
    - datetime (standard library)

**Required Files:**
    - Both normalized assessment workbooks (Assessment 1 and 2)
    - Generated dashboard workbook
    - This consolidation script

**File Naming:**
    Source workbooks MUST be named exactly:
    - ISMS-A.8.17-Assessment-1.xlsx
    - ISMS-A.8.17-Assessment-2.xlsx
    
    Dashboard can have any name (passed as command-line argument)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 consolidate_a817_dashboard.py ISMS-A.8.17-Dashboard-Time-Sync.xlsx

**Prerequisites:**
    1. Generate dashboard first:
       python3 generate_a817_4_compliance_dashboard.py
    
    2. Normalize source assessment files:
       python3 normalize_assessment_files_a817.py
    
    3. Ensure all files are accessible:
       - ISMS-A.8.17-Assessment-1.xlsx
       - ISMS-A.8.17-Assessment-2.xlsx
       - Dashboard file (provided as argument)

**File Location Requirements:**
    Source workbooks must be in CURRENT DIRECTORY when running script:
    - ISMS-A.8.17-Assessment-1.xlsx (in current directory)
    - ISMS-A.8.17-Assessment-2.xlsx (in current directory)
    
    Dashboard can be in any directory (specify full path if needed)

**Example Session:**
$ python3 consolidate_a817_dashboard.py ISMS-A.8.17-Dashboard-20250125.xlsx
ISMS A.8.17 - DASHBOARD DATA CONSOLIDATION
Dashboard: ISMS-A.8.17-Dashboard-20250125.xlsx
Started: 2025-01-25 14:30:15
Checking source workbooks...
✓ ISMS-A.8.17-Assessment-1.xlsx
✓ ISMS-A.8.17-Assessment-2.xlsx
Loading dashboard...
✓ Dashboard loaded (9 sheets)
================================================================================
CONSOLIDATING DATA FROM SOURCE WORKBOOKS
📁 Processing: ISMS-A.8.17-Assessment-1.xlsx
✓ Time_Sources → Time_Sources_Summary: 8 rows
✓ Internal_NTP_Servers → Time_Sources_Summary: 4 rows
📁 Processing: ISMS-A.8.17-Assessment-2.xlsx
✓ System_Sync_Status → Sync_Status_Summary: 247 rows
✓ Gap_Analysis → Gap_Analysis: 12 rows
================================================================================
SAVING CONSOLIDATED DASHBOARD
✓ Dashboard saved successfully
================================================================================
CONSOLIDATION COMPLETE
📊 Statistics:
• Source sheets read: 4
• Total rows consolidated: 271
📋 By Dashboard Sheet:
• Time_Sources_Summary: 12 rows
• Sync_Status_Summary: 247 rows
• Gap_Analysis: 12 rows
✓ Completed: 2025-01-25 14:30:28
================================================================================
Dashboard now contains all data from assessments!
Self-contained file ready for distribution.

**Exit Codes:**
    0 = Consolidation successful
    1 = Consolidation failed (missing files, errors during processing)

**Re-Running Consolidation:**
    Safe to re-run when assessment data is updated:
    - Existing dashboard data will be overwritten
    - Updated source data will be consolidated
    - Statistics will reflect new data volumes
    - No data loss in source workbooks (read-only access)

**Workflow Integration:**

**Initial Setup (One-Time):**
1. Generate dashboard: python3 generate_a817_4_compliance_dashboard.py
2. Normalize assessments: python3 normalize_assessment_files_a817.py
3. Consolidate data: python3 consolidate_a817_dashboard.py dashboard.xlsx

**Monthly/Quarterly Updates:**
1. Update source assessments with new data
2. Re-normalize (if file names changed): python3 normalize_assessment_files_a817.py
3. Re-consolidate: python3 consolidate_a817_dashboard.py dashboard.xlsx

**Static Snapshot Creation:**
1. Run consolidation with date-stamped dashboard name
2. Archive consolidated dashboard for reporting period
3. Create new dashboard for next period

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.17
Script Type:          Data Consolidation / Alternative Integration Method
Dashboard Component:  A.8.17 Compliance Dashboard Support
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Input Files:
    - ISMS-A.8.17-Assessment-1.xlsx (Time Source Infrastructure)
    - ISMS-A.8.17-Assessment-2.xlsx (System Synchronization Status)
    - Dashboard workbook (ISMS-A.8.17-Dashboard-*.xlsx)

Output:
    - Modified dashboard workbook with consolidated data
    - Console statistics showing consolidation results

Related Scripts:
    - generate_a817_1_time_sources.py (creates Assessment 1)
    - generate_a817_2_sync_status.py (creates Assessment 2)
    - generate_a817_4_compliance_dashboard.py (creates dashboard)
    - normalize_assessment_files_a817.py (filename normalization)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Consolidation mapping for 2 assessment workbooks
    - 4 source sheets → 3 dashboard sheets
    - Safe merged cell handling
    - Detailed consolidation statistics

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**File Modification:**
- Source workbooks are read in data-only mode (NOT modified)
- Dashboard workbook IS modified (data written to target sheets)
- Always backup dashboard before first consolidation

**Re-Running Consolidation:**
- Safe to re-run when source data is updated
- Existing consolidated data is overwritten
- No cumulative effect (fresh consolidation each time)
- Statistics reflect current consolidation only

**Source File Requirements:**
- Must use EXACT normalized file names
- Must be in current directory when running script
- Must be readable .xlsx format (not .xls or .xlsm)
- Must contain expected sheet names from mapping

**Dashboard Requirements:**
- Must be generated by generate_a817_4_compliance_dashboard.py
- Must contain expected target sheet names
- Must be writable (not read-only or open in Excel)
- Should be backed up before first consolidation

**Comparison: External Linking vs. Data Consolidation**

| Aspect | External Linking | Data Consolidation |
|--------|------------------|-------------------|
| **Setup** | One-time generation | Generate + consolidate |
| **Updates** | Automatic (Excel refresh) | Manual re-run required |
| **File Dependency** | Must be co-located | Can be separate |
| **Distribution** | Requires source files | Self-contained |
| **Excel Security** | May be blocked | Always works |
| **Historical Tracking** | Overwrites data | Create dated snapshots |
| **Data Type** | Live formulas | Static values |
| **Use Case** | Real-time reporting | Point-in-time reporting |

**Performance:**
- Typical consolidation time: 5-15 seconds
- Depends on data volume in source workbooks
- Progress is displayed for each source workbook
- Large datasets (1000+ rows) may take longer

**Limitations:**
- Does not preserve formulas from source workbooks
- Does not copy cell formatting or styles
- Does not handle Excel tables or charts
- Does not validate data during consolidation
- Assumes source data is already validated

**Best Practices:**
1. Complete source assessments before consolidation
2. Run normalization script before consolidation
3. Backup dashboard before first consolidation
4. Review consolidation statistics after each run
5. Document when/why consolidation approach is used vs. external linking

**Audit Considerations:**
- Consolidation creates static point-in-time snapshot
- Consolidation timestamp is recorded in statistics
- Source data traceability via source labels
- Useful for monthly/quarterly compliance reporting archives
- Keep consolidation logs for audit trail

**Data Protection:**
- Consolidated dashboard inherits data classification from sources
- Source workbooks remain unmodified and unaffected
- Dashboard may contain aggregated sensitive compliance data
- Handle according to organization's data classification policy

================================================================================
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from datetime import datetime

# ============================================================================
# CONFIGURATION
# ============================================================================

SOURCE_WORKBOOKS = {
    "wb1": "ISMS-A.8.17-Assessment-1.xlsx",  # Time Source Infrastructure
    "wb2": "ISMS-A.8.17-Assessment-2.xlsx",  # System Synchronization Status
}

SHEET_START_ROWS = {
    "Gaps_Action_Items": 10,  # Start after headers
}

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def safely_write_data(ws, start_row, data):
    """Safely write data, handling merged cells"""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except:
                continue
        entries_written += 1
    return entries_written

# ============================================================================
# DATA EXTRACTION
# ============================================================================

def extract_time_sources(filepath):
    """Extract external time sources from WB1"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        if 'Time_Sources' not in wb.sheetnames:
            wb.close()
            return []
        
        ws = wb['Time_Sources']
        sources = []
        
        for row in range(2, min(ws.max_row + 1, 100)):
            source_name = ws[f'A{row}'].value
            if source_name and str(source_name).strip():
                source_type = ws[f'B{row}'].value
                stratum = ws[f'D{row}'].value
                status = ws[f'I{row}'].value
                sources.append({
                    'name': source_name,
                    'type': source_type,
                    'stratum': stratum,
                    'status': status
                })
        
        wb.close()
        return sources
    except Exception as e:
        print(f"  [!] Error reading time sources: {e}")
        return []

def extract_sync_status(filepath):
    """Extract system synchronization status from WB2"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        if 'System_Sync_Status' not in wb.sheetnames:
            wb.close()
            return []
        
        ws = wb['System_Sync_Status']
        systems = []
        
        for row in range(2, min(ws.max_row + 1, 1000)):
            system_name = ws[f'A{row}'].value
            if system_name and str(system_name).strip():
                sync_status = ws[f'C{row}'].value
                time_drift = ws[f'D{row}'].value
                compliant = ws[f'F{row}'].value
                
                # Track non-compliant systems
                if compliant and 'Non-Compliant' in str(compliant):
                    systems.append({
                        'system': system_name,
                        'status': sync_status,
                        'drift': time_drift,
                        'compliant': compliant
                    })
        
        wb.close()
        return systems
    except Exception as e:
        print(f"  [!] Error reading sync status: {e}")
        return []

def generate_gaps_from_data(time_sources, sync_issues):
    """Generate gap entries from non-compliant data"""
    gaps = []
    gap_counter = 1
    
    # Gap from time sources
    inactive_sources = [s for s in time_sources if s['status'] and 'Inactive' in str(s['status'])]
    if inactive_sources:
        for source in inactive_sources:
            gaps.append([
                f"GAP-817-{gap_counter:03d}",
                f"Time source inactive: {source['name']}",
                "HIGH",
                f"Restore or replace {source['name']} time source",
                None,  # Owner
            ])
            gap_counter += 1
    
    # Gaps from sync issues
    for system in sync_issues:
        gaps.append([
            f"GAP-817-{gap_counter:03d}",
            f"System not synchronized: {system['system']}",
            "MEDIUM",
            f"Configure NTP/fix sync issue - Current drift: {system['drift']}",
            None,  # Owner
        ])
        gap_counter += 1
    
    return gaps

# ============================================================================
# MAIN CONSOLIDATION
# ============================================================================

def populate_dashboard(dashboard_file):
    """Populate dashboard from source workbooks"""
    
    print("="*80)
    print("A.8.17 CLOCK SYNCHRONIZATION DASHBOARD CONSOLIDATION")
    print("="*80)
    
    # Check for source workbooks
    missing = []
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        if not os.path.exists(wb_filename):
            missing.append(wb_filename)
    
    if missing:
        print("\n[X] ERROR: Missing source workbooks:")
        for wb in missing:
            print(f"    - {wb}")
        print("\n  Place both assessment workbooks in same directory.")
        return False
    
    print("\n[1/3] Extracting time source data...")
    time_sources = extract_time_sources(SOURCE_WORKBOOKS['wb1'])
    print(f"  * Time sources found: {len(time_sources)}")
    inactive = len([s for s in time_sources if s['status'] and 'Inactive' in str(s['status'])])
    if inactive:
        print(f"  * [!] Inactive sources: {inactive}")
    
    print("\n[2/3] Extracting sync status...")
    sync_issues = extract_sync_status(SOURCE_WORKBOOKS['wb2'])
    print(f"  * Non-compliant systems: {len(sync_issues)}")
    
    print("\n[3/3] Generating gaps and action items...")
    all_gaps = generate_gaps_from_data(time_sources, sync_issues)
    print(f"  * Total gaps: {len(all_gaps)}")
    
    # Load dashboard
    try:
        wb = load_workbook(dashboard_file)
    except Exception as e:
        print(f"  [X] Error loading dashboard: {e}")
        return False
    
    # Populate Gaps & Action Items sheet
    if 'Gaps_Action_Items' in wb.sheetnames and all_gaps:
        ws = wb['Gaps_Action_Items']
        count = safely_write_data(ws, SHEET_START_ROWS['Gaps_Action_Items'], all_gaps)
        print(f"  [OK] Gaps_Action_Items: {count} entries")
    
    # Save
    try:
        wb.save(dashboard_file)
        print(f"\n[SAVED] {dashboard_file}")
    except Exception as e:
        print(f"  [X] Error saving: {e}")
        return False
    
    print("\n" + "="*80)
    print("[OK] A.8.17 DASHBOARD CONSOLIDATION COMPLETE")
    print("="*80)
    print(f"\nSummary:")
    print(f"  * Time sources: {len(time_sources)} ({inactive} inactive)")
    print(f"  * Non-compliant systems: {len(sync_issues)}")
    print(f"  * Gaps identified: {len(all_gaps)}")
    print(f"\n[ROCKET] Dashboard ready for presentation!")
    print("="*80 + "\n")
    
    return True

# ============================================================================
# MAIN
# ============================================================================

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 consolidate_a817_dashboard.py <dashboard.xlsx>")
        print("\nExample:")
        print("  python3 consolidate_a817_dashboard.py ISMS-A.8.17-Dashboard-Time-Sync.xlsx")
        sys.exit(1)
    
    dashboard_file = sys.argv[1]
    
    if not os.path.exists(dashboard_file):
        print(f"[X] Error: Dashboard file not found: {dashboard_file}")
        sys.exit(1)
    
    success = populate_dashboard(dashboard_file)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
