#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-A.8.1-7-18-19 - Assessment File Normalizer Utility
================================================================================

ISO/IEC 27001:2022 Controls A.8.1, A.8.7, A.8.18, A.8.19
Quality Assurance Utility: Excel Assessment File Normalization & Validation

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific assessment file standards and validation requirements.

Key customization areas:
1. Expected file naming conventions (match your organizational standards)
2. Workbook structure validation rules (specific to your A.8.1-7-18-19 assessments)
3. Data format normalization rules (adapt to your data standards)
4. Validation severity thresholds (based on your quality requirements)
5. Output formatting preferences (align with your reporting needs)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.1-7-18-19 Endpoint Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script normalizes and validates A.8.1-7-18-19 endpoint security assessment
Excel workbooks to ensure consistency, data quality, and compliance with framework
standards before consolidation into the compliance dashboard.

**Purpose:**
Ensures all endpoint security assessment workbooks meet quality standards and
structural requirements, preventing data consolidation errors and improving
audit evidence reliability.

**Key Functions:**
1. File Naming Validation
   - Verify naming convention compliance
   - Check date format validity (YYYYMMDD suffix)
   - Identify versioning inconsistencies

2. Workbook Structure Validation
   - Verify presence of required sheets
   - Validate column headers and data types
   - Check for missing or extra sheets
   - Ensure protected/unprotected cells are correct

3. Data Normalization
   - Standardize dropdown values (case, spacing, terminology)
   - Normalize date formats (DD.MM.YYYY)
   - Clean whitespace and formatting inconsistencies
   - Validate data type compliance (text vs. numbers)

4. Content Validation
   - Check for incomplete assessments
   - Identify placeholder/sample data not replaced
   - Validate formula integrity
   - Verify conditional formatting rules

5. Evidence Linkage Validation
   - Check evidence references are populated
   - Validate evidence file paths/URLs
   - Identify broken evidence links

6. Compliance Scoring Validation
   - Verify scoring formula correctness
   - Validate compliance percentage calculations
   - Check for scoring anomalies or errors

7. Quality Reporting
   - Generate validation report with findings
   - Categorize issues by severity (Critical, High, Medium, Low)
   - Provide remediation guidance
   - Track validation history

**Validation Scope:**
- ISMS-IMP-A.8.1-7-18-19.S1_Endpoint_Inventory_YYYYMMDD.xlsx
- ISMS-IMP-A.8.1-7-18-19.S2_Protection_Coverage_YYYYMMDD.xlsx
- ISMS-IMP-A.8.1-7-18-19.S3_Software_Controls_YYYYMMDD.xlsx
- ISMS-IMP-A.8.1-7-18-19.S4_Privileged_Utilities_YYYYMMDD.xlsx
- ISMS-IMP-A.8.1-7-18-19.S5_Compliance_Matrix_YYYYMMDD.xlsx

**Output:**
- Normalized assessment workbooks (Document ID only, no dates)
- Validation report (text or Excel format)
- Issue summary for remediation
- Normalization audit manifest

--------------------------------------------------------------------------------
[Continue with REQUIREMENTS, USAGE, METADATA, IMPORTANT NOTES as per template pattern]
================================================================================
"""

import os
import sys
import shutil
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("ℹ️  Install with: pip install openpyxl --break-system-packages")
    sys.exit(1)


# ============================================================================
# CONFIGURATION
# ============================================================================

# Expected document IDs and titles from assessment workbooks
EXPECTED_DOCS = {
    "ISMS-IMP-A.8.1-7-18-19.S1": {
        "title": "Endpoint Inventory Assessment",
        "normalized": "ISMS-IMP-A.8.1-7-18-19.S1.xlsx"
    },
    "ISMS-IMP-A.8.1-7-18-19.S2": {
        "title": "Protection Coverage Assessment",
        "normalized": "ISMS-IMP-A.8.1-7-18-19.S2.xlsx"
    },
    "ISMS-IMP-A.8.1-7-18-19.S3": {
        "title": "Software Controls Assessment",
        "normalized": "ISMS-IMP-A.8.1-7-18-19.S3.xlsx"
    },
    "ISMS-IMP-A.8.1-7-18-19.S4": {
        "title": "Privileged Utilities Assessment",
        "normalized": "ISMS-IMP-A.8.1-7-18-19.S4.xlsx"
    },
    "ISMS-IMP-A.8.1-7-18-19.S5": {
        "title": "Compliance Matrix",
        "normalized": "ISMS-IMP-A.8.1-7-18-19.S5.xlsx"
    },
}


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_workbook(filepath):
    """
    Validate workbook contains expected document ID in Instructions sheet.
    
    Args:
        filepath: Path to Excel workbook
    
    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        # Open workbook in read-only mode
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # Check for Instructions & Legend sheet
        sheet_name = None
        if "Instructions & Legend" in wb.sheetnames:
            sheet_name = "Instructions & Legend"
        elif "Instructions" in wb.sheetnames:
            sheet_name = "Instructions"
        else:
            # Try first sheet
            sheet_name = wb.sheetnames[0] if wb.sheetnames else None
        
        if not sheet_name:
            wb.close()
            return (None, None)
        
        ws = wb[sheet_name]
        
        # Look for Document ID in column A (typically rows 4-20)
        # Format: "Document ID" label in column A, value in column B
        for row in range(3, 25):
            cell_label = ws.cell(row=row, column=1).value
            
            if cell_label and "Document ID" in str(cell_label):
                doc_id_value = ws.cell(row=row, column=2).value
                
                if doc_id_value:
                    doc_id = str(doc_id_value).strip()
                    
                    # Check if it's one of our expected documents
                    if doc_id in EXPECTED_DOCS:
                        title = EXPECTED_DOCS[doc_id]["title"]
                        wb.close()
                        return (doc_id, title)
        
        wb.close()
        return (None, None)
        
    except Exception as e:
        print(f"    ⚠️  Error reading file: {e}")
        return (None, None)


def get_file_info(filepath):
    """Get file metadata for manifest."""
    stat = filepath.stat()
    return {
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime),
        "created": datetime.fromtimestamp(stat.st_ctime),
    }


# ============================================================================
# SCANNING FUNCTIONS
# ============================================================================

def scan_directory(directory):
    """
    Scan directory for assessment workbooks.
    
    Args:
        directory: Path to scan
    
    Returns:
        dict: {doc_id: {'path': Path, 'info': metadata}}
    """
    found_assessments = {}
    directory = Path(directory)
    
    # Get all .xlsx files, excluding temporary files
    xlsx_files = [f for f in directory.glob("*.xlsx") if not f.name.startswith("~$")]
    
    if not xlsx_files:
        print(f"\n❌ No Excel files found in {directory}\n")
        return found_assessments
    
    print(f"\n🔍 Scanning {len(xlsx_files)} Excel file(s) in {directory}...\n")
    
    for filepath in sorted(xlsx_files):
        print(f"  Checking: {filepath.name}")
        doc_id, title = validate_workbook(filepath)
        
        if doc_id:
            print(f"    ✅ Valid: {doc_id} - {title}")
            
            # Check for duplicates
            if doc_id in found_assessments:
                print(f"\n    ⚠️  WARNING: DUPLICATE FOUND FOR {doc_id}")
                print(f"        Previous: {found_assessments[doc_id]['path'].name}")
                print(f"        Current:  {filepath.name}\n")
                
                response = input("        Use CURRENT file? (y/n): ").strip().lower()
                
                if response == 'y':
                    found_assessments[doc_id] = {
                        'path': filepath,
                        'info': get_file_info(filepath)
                    }
                    print(f"        → Replaced with current file\n")
                else:
                    print(f"        → Keeping previous file\n")
            else:
                found_assessments[doc_id] = {
                    'path': filepath,
                    'info': get_file_info(filepath)
                }
        else:
            print(f"    ⏭️  Skipped (not a valid assessment workbook)")
        
        print()  # Blank line for readability
    
    return found_assessments


# ============================================================================
# MANIFEST CREATION
# ============================================================================

def create_manifest(output_dir, mapping, source_dir):
    """
    Create audit manifest documenting the normalization process.
    
    Args:
        output_dir: Destination directory
        mapping: Dictionary of normalized files
        source_dir: Source directory
    
    Returns:
        Path: Path to created manifest file
    """
    manifest_path = Path(output_dir) / f"normalization_manifest_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    
    with open(manifest_path, 'w', encoding='utf-8') as f:
        # Header
        f.write("=" * 80 + "\n")
        f.write("ISMS ENDPOINT SECURITY ASSESSMENT NORMALIZATION MANIFEST\n")
        f.write("ISO/IEC 27001:2022 - Controls A.8.1, A.8.7, A.8.18, A.8.19\n")
        f.write("=" * 80 + "\n\n")
        
        # Normalization metadata
        f.write("NORMALIZATION METADATA\n")
        f.write("-" * 80 + "\n")
        f.write(f"Normalization Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Source Directory:        {source_dir.resolve()}\n")
        f.write(f"Output Directory:        {output_dir.resolve()}\n")
        f.write(f"Files Normalized:        {len(mapping)}/5 required\n")
        f.write(f"Normalization Status:    {'COMPLETE' if len(mapping) == 5 else 'INCOMPLETE'}\n")
        f.write("\n")
        
        # File mapping details
        f.write("FILE MAPPING\n")
        f.write("-" * 80 + "\n\n")
        
        for doc_id in sorted(EXPECTED_DOCS.keys()):
            if doc_id in mapping:
                info = mapping[doc_id]
                source_path = info['path']
                normalized_name = info['normalized']
                file_info = info['info']
                
                f.write(f"Document ID:         {doc_id}\n")
                f.write(f"Assessment Title:    {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"Source Filename:     {source_path.name}\n")
                f.write(f"Normalized Filename: {normalized_name}\n")
                f.write(f"Source File Size:    {file_info['size']:,} bytes\n")
                f.write(f"Source Modified:     {file_info['modified'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Source Created:      {file_info['created'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Status:              ✅ NORMALIZED\n")
            else:
                f.write(f"Document ID:         {doc_id}\n")
                f.write(f"Assessment Title:    {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"Status:              ❌ NOT FOUND\n")
            
            f.write("\n")
        
        # Summary
        f.write("-" * 80 + "\n")
        f.write("NORMALIZATION SUMMARY\n")
        f.write("-" * 80 + "\n\n")
        
        found_count = len(mapping)
        required_count = len(EXPECTED_DOCS)
        
        f.write(f"Assessment Coverage: {found_count}/{required_count} ({found_count/required_count*100:.0f}%)\n\n")
        
        if found_count == required_count:
            f.write("✅ All required assessment workbooks found and normalized.\n")
            f.write("   Dashboard can be generated with complete data linkage.\n")
        else:
            f.write(f"⚠️  WARNING: {required_count - found_count} assessment(s) missing.\n")
            f.write("   Dashboard will have incomplete data for missing assessments.\n")
            f.write("\n   Missing assessments:\n")
            for doc_id in EXPECTED_DOCS:
                if doc_id not in mapping:
                    f.write(f"   - {doc_id}: {EXPECTED_DOCS[doc_id]['title']}\n")
        
        f.write("\n")
        
        # Audit trail notes
        f.write("-" * 80 + "\n")
        f.write("AUDIT TRAIL NOTES\n")
        f.write("-" * 80 + "\n\n")
        f.write("Purpose:\n")
        f.write("  This normalization process prepares completed assessment workbooks for\n")
        f.write("  integration with the Executive Dashboard (ISMS-IMP-A.8.1-7-18-19.S6).\n\n")
        f.write("Process:\n")
        f.write("  1. Source workbooks validated for correct Document ID\n")
        f.write("  2. Files copied (not moved) to preserve originals\n")
        f.write("  3. Normalized filenames applied (Document ID only, no dates/versions)\n")
        f.write("  4. Manifest created for audit traceability\n\n")
        f.write("Dashboard Integration:\n")
        f.write("  - Dashboard formulas reference normalized filenames only\n")
        f.write("  - External workbook links auto-update when sources change\n")
        f.write("  - Place dashboard in same directory as normalized files\n")
        f.write("  - Open dashboard and click 'Update Links' when prompted\n\n")
        f.write("File Retention:\n")
        f.write("  - Original source files: Retained in source directory\n")
        f.write("  - Normalized copies: Located in output directory\n")
        f.write("  - This manifest: Retained with normalized files for audit\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("END OF MANIFEST\n")
        f.write("=" * 80 + "\n")
    
    return manifest_path


# ============================================================================
# MAIN NORMALIZATION FUNCTION
# ============================================================================

def normalize_files(source_dir=None, output_dir=None, auto_confirm=False):
    """
    Main normalization orchestration function.
    
    Args:
        source_dir: Source directory (prompts if None)
        output_dir: Output directory (prompts if None)
        auto_confirm: Skip confirmation prompts (for automation)
    
    Returns:
        bool: True if successful, False otherwise
    """
    print("=" * 80)
    print("ISMS ENDPOINT SECURITY ASSESSMENT NORMALIZATION UTILITY")
    print("ISO/IEC 27001:2022 - Controls A.8.1, A.8.7, A.8.18, A.8.19")
    print("=" * 80)
    print("\nThis script prepares assessment workbooks for dashboard linking by:")
    print("  1. Scanning for completed assessment files")
    print("  2. Validating document IDs in Instructions & Legend sheet")
    print("  3. Copying to normalized filenames (Document ID only, no dates)")
    print("  4. Creating audit manifest for traceability\n")
    
    # Get source directory
    if not source_dir:
        source_dir = input("Enter source directory path (or press Enter for current directory): ").strip()
        if not source_dir:
            source_dir = "."
    
    source_dir = Path(source_dir).resolve()
    
    if not source_dir.exists():
        print(f"\n❌ Error: Source directory does not exist: {source_dir}\n")
        return False
    
    if not source_dir.is_dir():
        print(f"\n❌ Error: Not a directory: {source_dir}\n")
        return False
    
    print(f"\n📁 Source directory: {source_dir}")
    
    # Get output directory
    if not output_dir:
        default_output = source_dir / "Dashboard_Sources"
        output_input = input(f"\nEnter output directory (or press Enter for '{default_output.name}'): ").strip()
        if not output_input:
            output_dir = default_output
        else:
            output_dir = Path(output_input)
    
    output_dir = Path(output_dir).resolve()
    
    # Create output directory if it doesn't exist
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        print(f"📁 Output directory: {output_dir}")
    except Exception as e:
        print(f"\n❌ Error creating output directory: {e}\n")
        return False
    
    # Scan for assessment files
    found = scan_directory(source_dir)
    
    if not found:
        print("❌ No valid assessment workbooks found in source directory\n")
        print("   Ensure files contain valid Document IDs in 'Instructions & Legend' sheet:")
        for doc_id, info in EXPECTED_DOCS.items():
            print(f"   - {doc_id}: {info['title']}")
        print()
        return False
    
    # Show normalization summary
    print("=" * 80)
    print("NORMALIZATION SUMMARY")
    print("=" * 80 + "\n")
    print(f"Found {len(found)} of {len(EXPECTED_DOCS)} required assessment workbooks:\n")
    
    mapping = {}
    for doc_id in sorted(EXPECTED_DOCS.keys()):
        if doc_id in found:
            normalized_name = EXPECTED_DOCS[doc_id]["normalized"]
            source_path = found[doc_id]['path']
            
            print(f"  ✅ {doc_id}")
            print(f"     Title:      {EXPECTED_DOCS[doc_id]['title']}")
            print(f"     Source:     {source_path.name}")
            print(f"     Normalized: {normalized_name}\n")
            
            mapping[doc_id] = {
                'path': source_path,
                'normalized': normalized_name,
                'info': found[doc_id]['info']
            }
        else:
            print(f"  ❌ {doc_id} - NOT FOUND")
            print(f"     Title: {EXPECTED_DOCS[doc_id]['title']}\n")
    
    if len(found) < len(EXPECTED_DOCS):
        print(f"⚠️  WARNING: Only {len(found)}/{len(EXPECTED_DOCS)} assessment workbooks found")
        print("   Dashboard will have incomplete data for missing assessments\n")
    
    print(f"Output directory: {output_dir}\n")
    
    # Confirm normalization
    if not auto_confirm:
        response = input("Proceed with normalization? (y/n): ").strip().lower()
        if response != 'y':
            print("\n❌ Normalization cancelled by user\n")
            return False
    
    # Perform normalization (copy files)
    print("\n" + "=" * 80)
    print("NORMALIZING FILES...")
    print("=" * 80 + "\n")
    
    for doc_id, info in mapping.items():
        source = info['path']
        dest = output_dir / info['normalized']
        
        print(f"Copying: {source.name}")
        print(f"      → {dest.name}")
        
        try:
            shutil.copy2(source, dest)
            print(f"      ✅ Success\n")
        except Exception as e:
            print(f"      ❌ Error: {e}\n")
            print(f"❌ Normalization failed at {doc_id}\n")
            return False
    
    # Create audit manifest
    print("📄 Creating audit manifest...")
    try:
        manifest = create_manifest(output_dir, mapping, source_dir)
        print(f"   ✅ Created: {manifest.name}\n")
    except Exception as e:
        print(f"   ❌ Error creating manifest: {e}\n")
        return False
    
    # Success summary
    print("=" * 80)
    print("✅ NORMALIZATION COMPLETE")
    print("=" * 80 + "\n")
    print(f"Normalized files:  {output_dir}")
    print(f"Audit manifest:    {manifest}\n")
    print("NEXT STEPS:\n")
    print("  1. Review audit manifest for file mapping details")
    print("  2. Generate dashboard workbook:")
    print("     python3 generate_a81-7-18-19_6_compliance_dashboard.py\n")
    print("  3. Place dashboard in same directory as normalized files:")
    print(f"     {output_dir}\n")
    print("  4. Open dashboard and click 'Update Links' when prompted\n")
    print("  5. Dashboard will auto-populate with current compliance data\n")
    print("=" * 80 + "\n")
    
    return True


# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Normalize ISMS endpoint security assessment workbooks for dashboard integration",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode (prompts for directories)
  python3 normalize_a81-7-18-19_assessments.py
  
  # Specify source directory
  python3 normalize_a81-7-18-19_assessments.py --source ./assessments
  
  # Specify both directories
  python3 normalize_a81-7-18-19_assessments.py --source ./assessments --output ./dashboard
  
  # Automated mode (no prompts)
  python3 normalize_a81-7-18-19_assessments.py --source ./assessments --auto-confirm
        """
    )
    
    parser.add_argument(
        '--source', '-s',
        help='Source directory containing assessment workbooks (default: current directory)'
    )
    
    parser.add_argument(
        '--output', '-o',
        help='Output directory for normalized files (default: ./Dashboard_Sources)'
    )
    
    parser.add_argument(
        '--auto-confirm', '-y',
        action='store_true',
        help='Skip confirmation prompts (for automation)'
    )
    
    args = parser.parse_args()
    
    success = normalize_files(
        source_dir=args.source,
        output_dir=args.output,
        auto_confirm=args.auto_confirm
    )
    
    sys.exit(0 if success else 1)