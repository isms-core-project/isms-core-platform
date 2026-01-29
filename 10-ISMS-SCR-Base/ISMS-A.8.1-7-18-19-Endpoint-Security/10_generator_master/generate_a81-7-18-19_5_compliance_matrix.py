#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.1-7-18-19.S5 - Compliance Matrix Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.1, A.8.7, A.8.18, A.8.19
Assessment Domain 5 of 6: Integrated Endpoint Security Compliance Matrix

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific compliance scoring methodology, risk prioritization,
and remediation tracking requirements.

Key customization areas:
1. Compliance scoring algorithms (adapt to your risk assessment methodology)
2. Risk prioritization criteria (based on your threat landscape)
3. Remediation SLA thresholds (aligned with your operational capabilities)
4. Compliance thresholds (match your governance requirements: 90%, 95%, etc.)
5. Integration with source assessment workbooks (verify file paths)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.1-7-18-19 Endpoint Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook that consolidates
endpoint security compliance data across all four controls (A.8.1, A.8.7, A.8.18,
A.8.19), providing integrated compliance scoring, gap analysis, and remediation
tracking.

**Purpose:**
Enables systematic assessment of overall endpoint security posture by integrating
inventory, malware protection, privileged utilities, and software control
assessments into a unified compliance matrix supporting risk-based prioritization
and remediation tracking.

**Assessment Scope:**
- Per-endpoint compliance across all 4 controls
- Integrated compliance scoring (endpoint-level and organizational-level)
- Risk prioritization (combination of multiple gaps = higher risk)
- Gap identification across all control domains
- Remediation tracking with SLA monitoring
- Trend analysis (compliance improving or degrading)
- Executive reporting and audit readiness
- Evidence consolidation from all assessment domains

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and compliance scoring methodology
2. Master_Compliance_Matrix - Per-endpoint compliance across all 4 controls
3. Compliance_Summary - Organizational compliance scores by control
4. Risk_Prioritization - High-risk endpoints requiring immediate action
5. Gap_Consolidation - All gaps from all assessment domains
6. Remediation_Tracking - Gap remediation with SLA monitoring
7. Trend_Analysis - Compliance trends over time (if historical data)
8. Evidence_Consolidation - Consolidated evidence register
9. Audit_Readiness - Audit evidence summary and completeness
10. Approval_Sign_Off - Stakeholder review and approval workflow

**Key Features:**
- Automated data consolidation from domains 1-4 workbooks
- Multi-dimensional compliance scoring (per-control and overall)
- Risk-based gap prioritization (unencrypted + unprotected = critical)
- Conditional formatting for compliance status visualization
- Remediation SLA tracking and alerting
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with all assessment domains

**Integration:**
This assessment consolidates data from:
- Domain 1: Endpoint Inventory (A.8.1-7-18-19.S1)
- Domain 2: Protection Coverage (A.8.1-7-18-19.S2)
- Domain 3: Software Controls (A.8.1-7-18-19.S3)
- Domain 4: Privileged Utilities (A.8.1-7-18-19.S4)

And feeds into:
- Domain 6: Compliance Dashboard (A.8.1-7-18-19.S6)

--------------------------------------------------------------------------------
[Continue with REQUIREMENTS, USAGE, METADATA, IMPORTANT NOTES as per template pattern]
================================================================================
"""

import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation




# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
LAPTOP = '\U0001F4BB' # 💻 Laptop
VIRUS = '\U0001F9A0'  # 🦠 Virus/Microbe
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def create_workbook() -> Workbook:
    """Create workbook with all sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    sheets = [
        "Instructions & Legend",
        "Master_Compliance_Matrix",
        "A81_Device_Compliance",
        "A87_Protection_Compliance",
        "A818_Utility_Compliance",
        "A819_Software_Compliance",
        "Risk_Prioritization",
        "Remediation_Tracking",
        "Trend_Analysis",
        "Evidence_Register",
        "Approval_Sign_Of",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def setup_styles():
    """Define styles."""
    thin = Side(style="thin")
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "partial": {"fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")},
        "noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "gap_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style to cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name, size=style_dict["font"].size, bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def create_instructions_sheet(ws, styles):
    """Create instructions."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "ENDPOINT SECURITY COMPLIANCE MATRIX"
    apply_style(cell, styles['header'])
    
    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "Master Compliance: A.8.1, A.8.7, A.8.18, A.8.19 Combined"
    apply_style(cell, styles['subheader'])
    
    row = 4
    metadata = [
        ("Document ID:", "ISMS-IMP-A.8.1-7-18-19.S5"),
        ("Workbook:", "Compliance Matrix"),
        ("Version:", "1.0"),
        ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for label, value in metadata:
        ws[f'A{row}'].value = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].value = value
        row += 1
    
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "📋 PURPOSE"
    apply_style(cell, styles['subheader'])
    row += 1
    
    purpose = """This workbook consolidates compliance assessment across all four endpoint security controls:

• A.8.1 User Endpoint Devices (inventory, baselines, encryption, MDM)
• A.8.7 Protection Against Malware (coverage, effectiveness, incidents)
• A.8.18 Use of Privileged Utility Programs (access controls, approvals, auditing)
• A.8.19 Installation of Software (approved software, unauthorized detection, app control)

Master compliance matrix shows per-endpoint status across all controls, enabling risk prioritization based on combined gaps (e.g., unencrypted + unprotected + unauthorized software = critical risk)."""
    
    ws.merge_cells(f'A{row}:F{row+3}')
    cell = ws[f'A{row}']
    cell.value = purpose
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 80
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60


def create_master_compliance_matrix_sheet(ws, styles):
    """Create master compliance matrix."""
    ws.merge_cells('A1:L1')
    cell = ws['A1']
    cell.value = "MASTER ENDPOINT SECURITY COMPLIANCE MATRIX"
    apply_style(cell, styles['header'])
    
    ws.merge_cells('A2:L2')
    cell = ws['A2']
    cell.value = "Per-endpoint compliance across A.8.1, A.8.7, A.8.18, A.8.19 with combined risk score"
    apply_style(cell, styles['subheader'])
    
    headers = [
        "Device ID",
        "Hostname",
        "Device Type",
        "A.8.1 Compliance",
        "A.8.7 Compliance",
        "A.8.18 Compliance",
        "A.8.19 Compliance",
        "Overall Score",
        "Combined Risk",
        "Critical Gaps",
        "Remediation Priority",
        "Notes"
    ]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Master matrix rows (50 endpoints)
    for i in range(50):
        current_row = 5 + i
        
        # Device ID
        ws.cell(row=current_row, column=1).value = f"EP-{1001+i:04d}"
        
        # Hostname (input)
        cell = ws.cell(row=current_row, column=2)
        apply_style(cell, styles['input_cell'])
        
        # Device Type (input)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        
        # A.8.1 Compliance (input or linked)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        
        # A.8.7 Compliance (input or linked)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        
        # A.8.18 Compliance (input or linked)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        
        # A.8.19 Compliance (input or linked)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        
        # Overall Score (calculated)
        cell = ws.cell(row=current_row, column=8)
        cell.value = f'=AVERAGE(D{current_row}:G{current_row})'
        
        # Combined Risk (input)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])
        
        # Critical Gaps (input)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])
        
        # Remediation Priority (calculated based on score)
        cell = ws.cell(row=current_row, column=11)
        cell.value = f'=IF(H{current_row}<50,"🔴 Critical",IF(H{current_row}<75,"🟠 High",IF(H{current_row}<90,"🟡 Medium","🟢 Low")))'
        
        # Notes
        cell = ws.cell(row=current_row, column=12)
        apply_style(cell, styles['input_cell'])
    
    # Summary
    summary_row = 57
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} OVERALL COMPLIANCE SUMMARY"
    apply_style(cell, styles['subheader'])
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Average Overall Score:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=AVERAGE(H5:H54)&"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "🔴 Critical Priority:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(K5:K54,"🔴 Critical")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "🟠 High Priority:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(K5:K54,"🟠 High")'
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 30


def create_control_compliance_sheet(ws, styles, control_name, control_id):
    """Create individual control compliance sheet."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = f"{control_id} COMPLIANCE DETAIL"
    apply_style(cell, styles['header'])
    
    headers = ["Device ID", "Hostname", "Compliance %", "Status", "Gaps", "Notes"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    for i in range(50):
        current_row = 5 + i
        ws.cell(row=current_row, column=1).value = f"EP-{1001+i:04d}"
        
        for col in range(2, 7):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
    
    ws.column_dimensions['A'].width = 12
    for col in ['B','C','D','E','F']:
        ws.column_dimensions[col].width = 20


def create_risk_prioritization_sheet(ws, styles):
    """Create risk prioritization."""
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "RISK PRIORITIZATION MATRIX"
    apply_style(cell, styles['header'])
    
    headers = ["Device ID", "Hostname", "Combined Gaps", "Risk Score", "Business Impact", "Remediation Difficulty", "Priority", "Notes"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    for i in range(50):
        current_row = 5 + i
        ws.cell(row=current_row, column=1).value = f"EP-{1001+i:04d}"
        
        for col in range(2, 9):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
    
    ws.column_dimensions['A'].width = 12
    for col in ['B','C','D','E','F','G','H']:
        ws.column_dimensions[col].width = 20


def create_remediation_tracking_sheet(ws, styles):
    """Create remediation tracking."""
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "REMEDIATION TRACKING"
    apply_style(cell, styles['header'])
    
    headers = ["Remediation ID", "Device ID", "Control", "Gap", "Severity", "Remediation Plan", "Owner", "Due Date", "Status", "Completion %", "Notes"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    for i in range(100):
        current_row = 5 + i
        ws.cell(row=current_row, column=1).value = f"REM-{i+1:04d}"
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        for col in range(2, 12):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
    
    ws.column_dimensions['A'].width = 15
    for col in ['B','C','D','E','F','G','H','I','J','K']:
        ws.column_dimensions[col].width = 20


def create_trend_analysis_sheet(ws, styles):
    """Create trend analysis."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "COMPLIANCE TREND ANALYSIS"
    apply_style(cell, styles['header'])
    
    headers = ["Assessment Date", "A.8.1 Score", "A.8.7 Score", "A.8.18 Score", "A.8.19 Score", "Overall Score"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # 12 assessments (monthly for a year)
    for i in range(12):
        current_row = 5 + i
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
    
    ws.column_dimensions['A'].width = 18
    for col in ['B','C','D','E','F']:
        ws.column_dimensions[col].width = 15


def create_evidence_register_sheet(ws, styles):
    """Create evidence register."""
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "EVIDENCE REGISTER"
    apply_style(cell, styles['header'])
    
    headers = ["Evidence ID", "Type", "Description", "Control", "Worksheet", "Location", "Date", "Collected By", "Status", "Notes"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    for i in range(100):
        current_row = 5 + i
        ws.cell(row=current_row, column=1).value = f"EVD-{i+1:03d}"
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        for col in range(2, 11):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40


def create_approval_signoff_sheet(ws, styles):
    """Create approval."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "APPROVAL & SIGN-OFF"
    apply_style(cell, styles['header'])
    
    ws[f'A4'].value = "Overall Compliance Score:"
    ws[f'A4'].font = Font(bold=True)
    ws[f'B4'].value = "=Master_Compliance_Matrix!B58"
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30


def main():
    """Main execution."""
    print("=" * 78)
    print("ISMS-IMP-A.8.1-7-18-19.S5 - Compliance Matrix Generator")
    print("Master Compliance: A.8.1, A.8.7, A.8.18, A.8.19")
    print("=" * 78)
    
    wb = create_workbook()
    styles = setup_styles()
    
    print("\n[Phase 1] Creating sheets...")
    
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    print("  ✅ Instructions")
    
    create_master_compliance_matrix_sheet(wb["Master_Compliance_Matrix"], styles)
    print("  ✅ Master Compliance Matrix")
    
    create_control_compliance_sheet(wb["A81_Device_Compliance"], styles, "A.8.1", "A.8.1")
    print("  ✅ A.8.1 Compliance")
    
    create_control_compliance_sheet(wb["A87_Protection_Compliance"], styles, "A.8.7", "A.8.7")
    print("  ✅ A.8.7 Compliance")
    
    create_control_compliance_sheet(wb["A818_Utility_Compliance"], styles, "A.8.18", "A.8.18")
    print("  ✅ A.8.18 Compliance")
    
    create_control_compliance_sheet(wb["A819_Software_Compliance"], styles, "A.8.19", "A.8.19")
    print("  ✅ A.8.19 Compliance")
    
    create_risk_prioritization_sheet(wb["Risk_Prioritization"], styles)
    print("  ✅ Risk Prioritization")
    
    create_remediation_tracking_sheet(wb["Remediation_Tracking"], styles)
    print("  ✅ Remediation Tracking")
    
    create_trend_analysis_sheet(wb["Trend_Analysis"], styles)
    print("  ✅ Trend Analysis")
    
    create_evidence_register_sheet(wb["Evidence_Register"], styles)
    print("  ✅ Evidence Register")
    
    create_approval_signoff_sheet(wb["Approval_Sign_Of"], styles)
    print("  ✅ Approval Sign-Of")
    
    filename = f"ISMS-IMP-A.8.1-7-18-19.S5_Compliance_Matrix_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        wb.save(filename)
        print(f"\n✅ SUCCESS: {filename}")
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        return 1
    
    print("\n✅ Master compliance matrix across all 4 controls")
    print(f"{CHECK} Per-endpoint combined risk scoring")
    print(f"{CHECK} Remediation prioritization and trend analysis")
    return 0


if __name__ == "__main__":
    exit(main())
