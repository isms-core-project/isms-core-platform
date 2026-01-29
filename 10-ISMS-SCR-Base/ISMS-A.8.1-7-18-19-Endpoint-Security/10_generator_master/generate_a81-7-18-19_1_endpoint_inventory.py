#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.1-7-18-19.S1 - Endpoint Inventory Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.1, A.8.7, A.8.18, A.8.19
Assessment Domain 1 of 6: Endpoint Device Inventory and Classification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific endpoint environment, discovery tools, and
inventory requirements.

Key customization areas:
1. Endpoint discovery methods (match your actual tools: MDM, SCCM, Jamf, etc.)
2. Classification categories (adapt to your device types and ownership models)
3. Security baseline requirements (specific to your endpoint standards)
4. Encryption technologies (based on your OS mix and encryption solutions)
5. Management platforms (aligned with your endpoint management infrastructure)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.1-7-18-19 Endpoint Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
endpoint device inventory, classification, security baseline compliance, and
management enrollment across all endpoint types.

**Purpose:**
Enables systematic assessment of endpoint inventory completeness and security
posture against ISO 27001:2022 Control A.8.1 requirements, supporting evidence-
based validation of endpoint device management and security controls.

**Assessment Scope:**
- Complete endpoint device inventory (laptops, desktops, mobile, tablets, IoT)
- Device classification (corporate-owned, BYOD, contractor, guest)
- Security baseline compliance per endpoint type
- Encryption status (full disk, file-level, mobile encryption)
- Management platform enrollment (MDM, agent-based, unmanaged)
- Remote wipe capabilities
- Lost/stolen device procedures
- Disposal and decommissioning tracking
- Endpoint access controls (authentication, screen lock)
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and classification framework
2. Inventory - Complete endpoint device inventory with 50 sample rows
3. Classification - Device type, ownership model, criticality analysis
4. Baseline_Compliance - Security baseline compliance per endpoint
5. Encryption_Status - Encryption technology and status tracking
6. Management_Enrollment - MDM/management platform enrollment
7. Capability_Requirements - Endpoint security capability assessment
8. Evidence_Register - Audit evidence tracking and documentation
9. Gap_Analysis - Non-compliant endpoints and remediation requirements
10. Approval_Sign_Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with dropdown lists for consistency
- Conditional formatting for compliance status visualization
- Automated gap identification for unmanaged/unencrypted devices
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with endpoint management tool outputs

**Integration:**
This assessment provides the inventory foundation for all other endpoint
security assessments (A.8.7 malware protection, A.8.18 privileged utilities,
A.8.19 software controls) and feeds into the A.8.1-7-18-19.S6 Compliance
Dashboard.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl --break-system-packages

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a81-7-18-19_1_endpoint_inventory.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a81-7-18-19_1_endpoint_inventory.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a81-7-18-19_1_endpoint_inventory.py --date 20250125

Output:
    File: ISMS-IMP-A.8.1-7-18-19.S1_Endpoint_Inventory_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize classification categories for your organization
    2. Inventory all endpoint devices using discovery tools
    3. Complete inventory sheet with actual endpoint data
    4. Classify endpoints by type, ownership, and criticality
    5. Assess baseline compliance for each endpoint
    6. Validate encryption status
    7. Verify management platform enrollment
    8. Conduct gap analysis for non-compliant endpoints
    9. Define remediation actions with timelines
    10. Collect and link audit evidence
    11. Obtain stakeholder approvals
    12. Feed results into other assessments and compliance dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.1 (Primary)
Assessment Domain:    1 of 6 (Endpoint Inventory and Classification)
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.1-7-18-19: Endpoint Security Framework (Governance)
    - ISMS-IMP-A.8.1-7-18-19-S1: Endpoint Discovery Process
    - ISMS-IMP-A.8.1-7-18-19-S2: Protection Coverage Assessment (Domain 2)
    - ISMS-IMP-A.8.1-7-18-19-S3: Software Controls Assessment (Domain 3)
    - ISMS-IMP-A.8.1-7-18-19-S4: Privileged Utilities Assessment (Domain 4)
    - ISMS-IMP-A.8.1-7-18-19-S5: Compliance Matrix (Domain 5)
    - ISMS-IMP-A.8.1-7-18-19-S6: Compliance Dashboard (Domain 6)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.1-7-18-19-S1 specification
    - Supports comprehensive endpoint inventory and classification
    - Integrated with A.8.1-7-18-19.S6 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Technology Diversity:**
Endpoint environments vary significantly across organizations. This script
provides a generic framework that must be customized for your specific:
- Endpoint types (Windows/Mac/Linux/mobile/IoT mix)
- Management platforms (Intune/Jamf/SCCM/MDM/etc.)
- Encryption solutions (BitLocker/FileVault/LUKS/mobile encryption)
- Ownership models (corporate/BYOD/contractor/guest policies)

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect complete endpoint inventory and classification.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Complete endpoint inventory with serial numbers and users
- Device locations and classifications
- Security posture and vulnerability information
- Organizational structure and access patterns

Handle in accordance with your organization's data classification policies.

**BYOD Considerations:**
Personal devices require different security controls and privacy protections.
Clearly distinguish BYOD from corporate-owned devices. Ensure BYOD policies
address data separation, limited management scope, and employee privacy rights.

**Maintenance:**
Review and update assessment:
- Monthly: Update inventory for new/decommissioned devices
- Quarterly: Refresh classification and baseline compliance
- Semi-annually: Comprehensive reassessment
- Annually: Full endpoint security review
- Ad-hoc: When significant infrastructure changes occur

**Quality Assurance:**
Have endpoint management team and security engineers validate inventory
completeness before using results for compliance reporting or remediation
decisions. Missing endpoints = missing security coverage.

**Continuous Improvement:**
Use assessment findings to improve endpoint management processes:
- Automate endpoint discovery and enrollment
- Implement continuous compliance monitoring
- Reduce time-to-remediation for gaps
- Improve baseline enforcement mechanisms

================================================================================
"""

import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================



# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
LAPTOP = '\U0001F4BB' # 💻 Laptop
VIRUS = '\U0001F9A0'  # 🦠 Virus/Microbe
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure for Endpoint Inventory Assessment
    sheets = [
        "Instructions & Legend",
        "Inventory",
        "Classification",
        "Baseline_Compliance",
        "Encryption_Status",
        "Management_Enrollment",
        "Capability_Requirements",
        "Evidence_Register",
        "Gap_Analysis",
        "Approval_Sign_Of",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        },
        "status_noncompliant": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_unknown": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        },
        "gap_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "gap_high": {
            "fill": PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),
        },
        "gap_medium": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
        },
        "gap_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    These are added to worksheet once, then applied to multiple cells.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_unknown': DataValidation(
            type="list",
            formula1='"Yes,No,Unknown"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Compliant,⚠️ Partial,❌ Non-Compliant,❓ Unknown"',
            allow_blank=False
        ),
        'device_type': DataValidation(
            type="list",
            formula1=f'"{LAPTOP} Laptop,🖥️ Desktop,📱 Mobile (iOS),📱 Mobile (Android),⌚ Tablet,🔧 Workstation,Other"',
            allow_blank=False
        ),
        'operating_system': DataValidation(
            type="list",
            formula1='"Windows 11,Windows 10,macOS 14 (Sonoma),macOS 13 (Ventura),iOS 17,iOS 16,Android 14,Android 13,Linux (Ubuntu),Linux (RHEL/CentOS),Other"',
            allow_blank=False
        ),
        'ownership_model': DataValidation(
            type="list",
            formula1='"🏢 Corporate-Owned,👤 BYOD,🤝 Contractor,👥 Shared Device,Other"',
            allow_blank=False
        ),
        'criticality': DataValidation(
            type="list",
            formula1='"🔴 Critical,🟠 High,🟡 Medium,🟢 Low"',
            allow_blank=False
        ),
        'encryption_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Encrypted,❌ Not Encrypted,⚠️ Partial,❓ Unknown"',
            allow_blank=False
        ),
        'encryption_technology': DataValidation(
            type="list",
            formula1='"BitLocker,FileVault,LUKS,Built-in (iOS/Android),Third-Party,None,Unknown"',
            allow_blank=False
        ),
        'mdm_platform': DataValidation(
            type="list",
            formula1='"Microsoft Intune,Jamf Pro,SCCM,VMware Workspace ONE,MobileIron,Citrix Endpoint Management,Kandji,Google Workspace,None,Other"',
            allow_blank=False
        ),
        'enrollment_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Enrolled,❌ Not Enrolled,⚠️ Partial,🔄 Pending,❓ Unknown"',
            allow_blank=False
        ),
        'baseline_compliance': DataValidation(
            type="list",
            formula1=f'"{CHECK} Compliant,⚠️ Partial,❌ Non-Compliant,🔄 Remediation in Progress,❓ Unknown"',
            allow_blank=False
        ),
        'firewall_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Enabled,❌ Disabled,⚠️ Partial,❓ Unknown"',
            allow_blank=False
        ),
        'antivirus_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Active,⚠️ Outdated,❌ Not Installed,🔴 Inactive,❓ Unknown"',
            allow_blank=False
        ),
        'device_location': DataValidation(
            type="list",
            formula1='"🏢 Office,🏠 Remote,✈️ Mobile,🌍 International,❓ Unknown"',
            allow_blank=False
        ),
        'device_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Active,⏸️ Inactive,🔒 Retired,📦 Storage,🔄 Pending Setup,❌ Lost/Stolen,❓ Unknown"',
            allow_blank=False
        ),
        'gap_severity': DataValidation(
            type="list",
            formula1='"🔴 Critical,🟠 High,🟡 Medium,🟢 Low"',
            allow_blank=False
        ),
        'gap_status': DataValidation(
            type="list",
            formula1='"🔴 Open,🟡 In Progress,🟢 Resolved,✅ Closed,⏸️ On Hold"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"📄 Config Export,📸 Screenshot,📊 Report,📜 Certificate,📋 Policy,📁 Log File,🔐 Encryption Key Backup,Other"',
            allow_blank=False
        ),
        'verification_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Verified,⏳ Pending,❌ Not Verified,⚠️ Needs Review"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1=f'"{CHECK} Approved,✅ Approved with Conditions,❌ Rejected,⏳ Pending Review"',
            allow_blank=False
        ),
    }

    # Add all validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)

    return validations


# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create comprehensive instructions and legend sheet."""
    
    # Document header
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "ENDPOINT INVENTORY & SECURITY ASSESSMENT"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "ISO/IEC 27001:2022 - Controls A.8.1, A.8.7, A.8.18, A.8.19"
    apply_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 20

    # Document metadata
    row = 4
    metadata = [
        ("Document ID:", "ISMS-IMP-A.8.1-7-18-19.S1"),
        ("Workbook:", "Endpoint Inventory Assessment"),
        ("Version:", "1.0"),
        ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Assessment Period:", "[To be completed by assessor]"),
        ("Assessor:", "[Name]"),
    ]
    
    for label, value in metadata:
        ws[f'A{row}'].value = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].value = value
        row += 1

    # Purpose section
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "📋 PURPOSE"
    apply_style(cell, styles['subheader'])
    row += 1

    purpose_text = """This workbook provides a comprehensive assessment of endpoint devices (laptops, desktops, mobile devices) for compliance with ISO/IEC 27001:2022 Controls A.8.1 (User Endpoint Devices), A.8.7 (Protection Against Malware), A.8.18 (Use of Privileged Utility Programs), and A.8.19 (Installation of Software on Operational Systems).

The assessment evaluates:
• Complete endpoint inventory and classification
• Security baseline compliance (OS hardening, firewall, encryption)
• Encryption coverage and key management
• MDM/endpoint management enrollment and capabilities
• Evidence collection and gap identification"""

    ws.merge_cells(f'A{row}:F{row+6}')
    cell = ws[f'A{row}']
    cell.value = purpose_text
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 110
    row += 7

    # Workbook structure
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = f"{CHART} WORKBOOK STRUCTURE"
    apply_style(cell, styles['subheader'])
    row += 1

    sheets_info = [
        ("Inventory", "Complete endpoint device inventory (Device ID, Hostname, Type, OS, User, Location, Status)"),
        ("Classification", "Device type distribution analysis and ownership model breakdown"),
        ("Baseline_Compliance", "Per-endpoint baseline compliance assessment (OS hardening, firewall, antivirus)"),
        ("Encryption_Status", "Encryption technology, status, key management per endpoint"),
        ("Management_Enrollment", "MDM/endpoint management platform, enrollment status, capabilities"),
        ("Capability_Requirements", "Policy requirements mapped to implementation (A.8.1 control mapping)"),
        ("Evidence_Register", "Comprehensive evidence documentation (100 evidence entries)"),
        ("Gap_Analysis", "Gap identification, severity classification, remediation tracking"),
        ("Approval_Sign_Of", "Multi-level approval workflow (Assessor → IT Ops Manager → CISO)"),
    ]

    headers = ["Sheet Name", "Description"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    for sheet_name, description in sheets_info:
        ws.cell(row=row, column=1).value = sheet_name
        ws.cell(row=row, column=2).value = description
        thin = Side(style="thin")
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

    # Status legend
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "🎨 STATUS LEGEND"
    apply_style(cell, styles['subheader'])
    row += 1

    legend_items = [
        (f"{CHECK} Compliant / Active / Enrolled", "Device meets all requirements, fully compliant", "status_compliant"),
        (f"{WARNING} Partial Compliance", "Device partially compliant, some requirements not met", "status_partial"),
        (f"{XMARK} Non-Compliant / Not Installed", "Device does not meet requirements, critical gap", "status_noncompliant"),
        ("❓ Unknown", "Status cannot be determined, requires investigation", "status_unknown"),
        ("🔴 Critical Severity", "Critical gap requiring immediate remediation (7 days)", "gap_critical"),
        ("🟠 High Severity", "High-priority gap requiring remediation (30 days)", "gap_high"),
        ("🟡 Medium Severity", "Medium-priority gap requiring remediation (60 days)", "gap_medium"),
        ("🟢 Low Severity", "Low-priority gap, scheduled remediation (90 days)", "gap_low"),
    ]

    headers = ["Status/Severity", "Meaning", "Color"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    for status, meaning, style_key in legend_items:
        ws.cell(row=row, column=1).value = status
        ws.cell(row=row, column=2).value = meaning
        cell = ws.cell(row=row, column=3)
        cell.value = "█████"
        if style_key in styles:
            apply_style(cell, styles[style_key])
        thin = Side(style="thin")
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

    # Usage instructions
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "📖 USAGE INSTRUCTIONS"
    apply_style(cell, styles['subheader'])
    row += 1

    instructions = [
        ("1. Inventory Sheet", "Complete endpoint inventory from MDM/SCCM/Jamf exports. Populate Device ID, Hostname, Type, OS, User, Location, Status for ALL endpoints."),
        ("2. Classification Sheet", "Auto-calculates device type distribution. Review for accuracy."),
        ("3. Baseline_Compliance", "Assess each endpoint against security baseline (OS hardening, firewall, encryption, antivirus). Use compliance status dropdowns."),
        ("4. Encryption_Status", "Document encryption technology per endpoint (BitLocker, FileVault, LUKS, built-in). Verify encryption active."),
        ("5. Management_Enrollment", "Document MDM platform enrollment status. Identify unmanaged devices."),
        ("6. Capability_Requirements", "Map policy requirements to implementation. Verify all 30 requirements addressed."),
        ("7. Evidence_Register", "Document ALL evidence (config exports, screenshots, reports, certificates). Minimum 20 evidence entries."),
        ("8. Gap_Analysis", "Document identified gaps with severity, root cause, remediation plan. Track remediation progress."),
        ("9. Approval_Sign_Of", "Obtain approvals: Assessor → IT Operations Manager → CISO. Final approval required for assessment closure."),
    ]

    headers = ["Step", "Instructions"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    for step, instruction in instructions:
        ws.cell(row=row, column=1).value = step
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = instruction
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        thin = Side(style="thin")
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.row_dimensions[row].height = 30
        row += 1

    # Important notes
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = f"{WARNING} IMPORTANT NOTES"
    apply_style(cell, styles['subheader'])
    row += 1

    notes = [
        f"{BULLET} This assessment is TECHNOLOGY-AGNOSTIC - works with ANY endpoint management platform (Intune, Jamf, SCCM, etc.)",
        f"{BULLET} Yellow-highlighted cells require user input - complete ALL yellow cells",
        f"{BULLET} Document ID must NOT be changed - used for workbook linking in dashboard",
        f"{BULLET} Evidence Register must contain minimum 20 evidence entries for audit compliance",
        f"{BULLET} Gap remediation SLAs: Critical (7 days), High (30 days), Medium (60 days), Low (90 days)",
        f"{BULLET} Obtain final CISO approval in Approval_Sign_Off sheet before submitting for audit",
    ]

    for note in notes:
        ws.cell(row=row, column=1).value = note
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15


# ============================================================================
# SECTION 4: INVENTORY SHEET
# ============================================================================

def create_inventory_sheet(ws, styles):
    """Create main endpoint inventory sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:P1')
    cell = ws['A1']
    cell.value = "ENDPOINT DEVICE INVENTORY"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:P2')
    cell = ws['A2']
    cell.value = "Complete inventory of all endpoint devices (laptops, desktops, mobile devices, tablets)"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Device ID",
        "Hostname",
        "Device Type",
        "Operating System",
        "OS Version",
        "Primary User",
        "Department",
        "Ownership Model",
        "Location",
        "Serial Number",
        "Asset Tag",
        "Purchase Date",
        "Last Seen",
        "Device Status",
        "Criticality",
        "Notes"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Sample data rows (50 rows for data entry)
    start_row = 5
    for i in range(50):
        current_row = start_row + i
        
        # Device ID (auto-generated format)
        ws.cell(row=current_row, column=1).value = f"EP-{1001+i:04d}"
        
        # Hostname (input cell)
        cell = ws.cell(row=current_row, column=2)
        apply_style(cell, styles['input_cell'])
        
        # Device Type (dropdown)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        validations['device_type'].add(cell)
        
        # Operating System (dropdown)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        validations['operating_system'].add(cell)
        
        # OS Version (input)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        
        # Primary User (input)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        
        # Department (input)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        
        # Ownership Model (dropdown)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])
        validations['ownership_model'].add(cell)
        
        # Location (dropdown)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])
        validations['device_location'].add(cell)
        
        # Serial Number (input)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])
        
        # Asset Tag (input)
        cell = ws.cell(row=current_row, column=11)
        apply_style(cell, styles['input_cell'])
        
        # Purchase Date (input)
        cell = ws.cell(row=current_row, column=12)
        apply_style(cell, styles['input_cell'])
        
        # Last Seen (input)
        cell = ws.cell(row=current_row, column=13)
        apply_style(cell, styles['input_cell'])
        
        # Device Status (dropdown)
        cell = ws.cell(row=current_row, column=14)
        apply_style(cell, styles['input_cell'])
        validations['device_status'].add(cell)
        
        # Criticality (dropdown)
        cell = ws.cell(row=current_row, column=15)
        apply_style(cell, styles['input_cell'])
        validations['criticality'].add(cell)
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=16)
        apply_style(cell, styles['input_cell'])

    # Summary statistics
    summary_row = start_row + 52
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} INVENTORY SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Total Endpoints:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTA(B5:B54)'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Active Devices:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(N5:N54,f"{CHECK} Active")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Inactive/Retired:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(N5:N54,"⏸️ Inactive")+COUNTIF(N5:N54,f"{LOCK} Retired")'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 30


# ============================================================================
# SECTION 5: CLASSIFICATION SHEET
# ============================================================================

def create_classification_sheet(ws, styles):
    """Create device classification and distribution analysis sheet."""
    
    # Header
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "ENDPOINT CLASSIFICATION & DISTRIBUTION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "Device type distribution, ownership model breakdown, and criticality analysis"
    apply_style(cell, styles['subheader'])

    # Device Type Distribution
    row = 4
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = f"{LAPTOP} DEVICE TYPE DISTRIBUTION"
    apply_style(cell, styles['subheader'])
    row += 1

    device_types = [
        (f"{LAPTOP} Laptop", "Inventory!C:C", f"{LAPTOP} Laptop"),
        ("🖥️ Desktop", "Inventory!C:C", "🖥️ Desktop"),
        ("📱 Mobile (iOS)", "Inventory!C:C", "📱 Mobile (iOS)"),
        ("📱 Mobile (Android)", "Inventory!C:C", "📱 Mobile (Android)"),
        ("⌚ Tablet", "Inventory!C:C", "⌚ Tablet"),
        ("🔧 Workstation", "Inventory!C:C", "🔧 Workstation"),
        ("Other", "Inventory!C:C", "Other"),
    ]

    headers = ["Device Type", "Count", "Percentage"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    total_row_start = row
    for device_type, range_ref, criteria in device_types:
        ws.cell(row=row, column=1).value = device_type
        ws.cell(row=row, column=2).value = f'=COUNTIF({range_ref},"{criteria}")'
        ws.cell(row=row, column=3).value = f'=IF(B{row+6}>0,B{row}/B{row+6}*100,0)&"%"'
        thin = Side(style="thin")
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

    # Total row
    ws.cell(row=row, column=1).value = "TOTAL"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = f'=SUM(B{total_row_start}:B{row-1})'
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 2

    # Ownership Model Distribution
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = "🏢 OWNERSHIP MODEL DISTRIBUTION"
    apply_style(cell, styles['subheader'])
    row += 1

    ownership_models = [
        ("🏢 Corporate-Owned", "Inventory!H:H", "🏢 Corporate-Owned"),
        ("👤 BYOD", "Inventory!H:H", "👤 BYOD"),
        ("🤝 Contractor", "Inventory!H:H", "🤝 Contractor"),
        ("👥 Shared Device", "Inventory!H:H", "👥 Shared Device"),
        ("Other", "Inventory!H:H", "Other"),
    ]

    headers = ["Ownership Model", "Count", "Percentage"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    total_row_start = row
    for model, range_ref, criteria in ownership_models:
        ws.cell(row=row, column=1).value = model
        ws.cell(row=row, column=2).value = f'=COUNTIF({range_ref},"{criteria}")'
        ws.cell(row=row, column=3).value = f'=IF(B{row+4}>0,B{row}/B{row+4}*100,0)&"%"'
        thin = Side(style="thin")
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

    # Total row
    ws.cell(row=row, column=1).value = "TOTAL"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = f'=SUM(B{total_row_start}:B{row-1})'
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 2

    # Criticality Distribution
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = f"{TARGET} CRITICALITY DISTRIBUTION"
    apply_style(cell, styles['subheader'])
    row += 1

    criticality_levels = [
        ("🔴 Critical", "Inventory!O:O", "🔴 Critical"),
        ("🟠 High", "Inventory!O:O", "🟠 High"),
        ("🟡 Medium", "Inventory!O:O", "🟡 Medium"),
        ("🟢 Low", "Inventory!O:O", "🟢 Low"),
    ]

    headers = ["Criticality Level", "Count", "Percentage"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    total_row_start = row
    for level, range_ref, criteria in criticality_levels:
        ws.cell(row=row, column=1).value = level
        ws.cell(row=row, column=2).value = f'=COUNTIF({range_ref},"{criteria}")'
        ws.cell(row=row, column=3).value = f'=IF(B{row+3}>0,B{row}/B{row+3}*100,0)&"%"'
        thin = Side(style="thin")
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

    # Total row
    ws.cell(row=row, column=1).value = "TOTAL"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = f'=SUM(B{total_row_start}:B{row-1})'
    ws.cell(row=row, column=2).font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15


# ============================================================================
# SECTION 6: BASELINE COMPLIANCE SHEET
# ============================================================================

def create_baseline_compliance_sheet(ws, styles):
    """Create per-endpoint baseline compliance assessment sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:L1')
    cell = ws['A1']
    cell.value = "ENDPOINT BASELINE COMPLIANCE ASSESSMENT"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:L2')
    cell = ws['A2']
    cell.value = "Per-endpoint security baseline compliance (OS hardening, firewall, antivirus, configuration)"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Device ID",
        "Hostname",
        "Device Type",
        "OS Hardening",
        "Firewall Status",
        "Antivirus Status",
        "Screen Lock",
        "Auto-Lock Time",
        "Password Policy",
        "Baseline Profile",
        "Compliance Status",
        "Notes"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Data rows (50 rows, linked to Inventory)
    start_row = 5
    for i in range(50):
        current_row = start_row + i
        
        # Device ID (linked to Inventory)
        ws.cell(row=current_row, column=1).value = f'=Inventory!A{5+i}'
        
        # Hostname (linked to Inventory)
        ws.cell(row=current_row, column=2).value = f'=Inventory!B{5+i}'
        
        # Device Type (linked to Inventory)
        ws.cell(row=current_row, column=3).value = f'=Inventory!C{5+i}'
        
        # OS Hardening (dropdown)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        validations['baseline_compliance'].add(cell)
        
        # Firewall Status (dropdown)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        validations['firewall_status'].add(cell)
        
        # Antivirus Status (dropdown)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        validations['antivirus_status'].add(cell)
        
        # Screen Lock (dropdown)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        validations['yes_no_unknown'].add(cell)
        
        # Auto-Lock Time (input)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])
        
        # Password Policy (dropdown)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])
        validations['baseline_compliance'].add(cell)
        
        # Baseline Profile (input)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])
        
        # Compliance Status (calculated based on other columns)
        cell = ws.cell(row=current_row, column=11)
        # Formula: If all checks are compliant, show compliant, else show status
        cell.value = f'=IF(AND(D{current_row}=f"{CHECK} Compliant",E{current_row}=f"{CHECK} Enabled",F{current_row}=f"{CHECK} Active",G{current_row}="Yes",I{current_row}=f"{CHECK} Compliant"),f"{CHECK} Compliant",IF(OR(D{current_row}=f"{XMARK} Non-Compliant",E{current_row}=f"{XMARK} Disabled",F{current_row}=f"{XMARK} Not Installed"),f"{XMARK} Non-Compliant",f"{WARNING} Partial"))'
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=12)
        apply_style(cell, styles['input_cell'])

    # Summary statistics
    summary_row = start_row + 52
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} BASELINE COMPLIANCE SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = f"{CHECK} Fully Compliant:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K5:K54,f"{CHECK} Compliant")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = f"{WARNING} Partially Compliant:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K5:K54,f"{WARNING} Partial")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = f"{XMARK} Non-Compliant:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K5:K54,f"{XMARK} Non-Compliant")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Compliance Rate:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=IF(COUNTA(K5:K54)>0,COUNTIF(K5:K54,f"{CHECK} Compliant")/COUNTA(K5:K54)*100,0)&"%"'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 30


# ============================================================================
# SECTION 7: ENCRYPTION STATUS SHEET
# ============================================================================

def create_encryption_status_sheet(ws, styles):
    """Create encryption status and key management sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "ENDPOINT ENCRYPTION STATUS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "Encryption technology, status, key management, and recovery procedures per endpoint"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Device ID",
        "Hostname",
        "Device Type",
        "Encryption Technology",
        "Encryption Status",
        "Full Disk Encryption",
        "Key Escrow/Recovery",
        "Recovery Key Location",
        "Last Verified Date",
        "Notes"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Data rows (50 rows, linked to Inventory)
    start_row = 5
    for i in range(50):
        current_row = start_row + i
        
        # Device ID (linked to Inventory)
        ws.cell(row=current_row, column=1).value = f'=Inventory!A{5+i}'
        
        # Hostname (linked to Inventory)
        ws.cell(row=current_row, column=2).value = f'=Inventory!B{5+i}'
        
        # Device Type (linked to Inventory)
        ws.cell(row=current_row, column=3).value = f'=Inventory!C{5+i}'
        
        # Encryption Technology (dropdown)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        validations['encryption_technology'].add(cell)
        
        # Encryption Status (dropdown)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        validations['encryption_status'].add(cell)
        
        # Full Disk Encryption (dropdown)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        validations['yes_no_unknown'].add(cell)
        
        # Key Escrow/Recovery (dropdown)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        validations['yes_no_na'].add(cell)
        
        # Recovery Key Location (input)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])
        
        # Last Verified Date (input)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])

    # Summary statistics
    summary_row = start_row + 52
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} ENCRYPTION COVERAGE SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = f"{CHECK} Encrypted:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E5:E54,f"{CHECK} Encrypted")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = f"{XMARK} Not Encrypted:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E5:E54,f"{XMARK} Not Encrypted")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Encryption Coverage Rate:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=IF(COUNTA(E5:E54)>0,COUNTIF(E5:E54,f"{CHECK} Encrypted")/COUNTA(E5:E54)*100,0)&"%"'
    
    summary_row += 2
    ws[f'A{summary_row}'].value = "BitLocker (Windows):"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D5:D54,"BitLocker")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "FileVault (macOS):"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D5:D54,"FileVault")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Built-in (iOS/Android):"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D5:D54,"Built-in (iOS/Android)")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "LUKS (Linux):"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D5:D54,"LUKS")'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 30


# ============================================================================
# SECTION 8: MANAGEMENT ENROLLMENT SHEET
# ============================================================================

def create_management_enrollment_sheet(ws, styles):
    """Create MDM/endpoint management enrollment sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "ENDPOINT MANAGEMENT ENROLLMENT"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:K2')
    cell = ws['A2']
    cell.value = "MDM platform enrollment status, capabilities, and management coverage"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Device ID",
        "Hostname",
        "Device Type",
        "MDM Platform",
        "Enrollment Status",
        "Management Agent Version",
        "Last Check-In",
        "Remote Wipe Capable",
        "Policy Enforcement",
        "Compliance Policy Applied",
        "Notes"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Data rows (50 rows, linked to Inventory)
    start_row = 5
    for i in range(50):
        current_row = start_row + i
        
        # Device ID (linked to Inventory)
        ws.cell(row=current_row, column=1).value = f'=Inventory!A{5+i}'
        
        # Hostname (linked to Inventory)
        ws.cell(row=current_row, column=2).value = f'=Inventory!B{5+i}'
        
        # Device Type (linked to Inventory)
        ws.cell(row=current_row, column=3).value = f'=Inventory!C{5+i}'
        
        # MDM Platform (dropdown)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        validations['mdm_platform'].add(cell)
        
        # Enrollment Status (dropdown)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        validations['enrollment_status'].add(cell)
        
        # Management Agent Version (input)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        
        # Last Check-In (input)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        
        # Remote Wipe Capable (dropdown)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])
        validations['yes_no_na'].add(cell)
        
        # Policy Enforcement (dropdown)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])
        validations['yes_no_unknown'].add(cell)
        
        # Compliance Policy Applied (input)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=11)
        apply_style(cell, styles['input_cell'])

    # Summary statistics
    summary_row = start_row + 52
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} MANAGEMENT ENROLLMENT SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = f"{CHECK} Enrolled:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E5:E54,f"{CHECK} Enrolled")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = f"{XMARK} Not Enrolled:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E5:E54,f"{XMARK} Not Enrolled")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Enrollment Rate:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=IF(COUNTA(E5:E54)>0,COUNTIF(E5:E54,f"{CHECK} Enrolled")/COUNTA(E5:E54)*100,0)&"%"'
    
    summary_row += 2
    ws[f'A{summary_row}'].value = "By MDM Platform:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Microsoft Intune:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D5:D54,"Microsoft Intune")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Jamf Pro:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D5:D54,"Jamf Pro")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "SCCM:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D5:D54,"SCCM")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "None:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(D5:D54,"None")'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 30


# ============================================================================
# SECTION 9: CAPABILITY REQUIREMENTS SHEET
# ============================================================================

def create_capability_requirements_sheet(ws, styles):
    """Create policy requirements mapping sheet (A.8.1 requirements)."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "CAPABILITY REQUIREMENTS MAPPING"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "A.8.1 Policy Requirements → Implementation Verification (30 requirements)"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Req ID",
        "Policy Requirement",
        "Implemented",
        "Evidence Reference",
        "Notes",
        "Status"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Requirements (from POL-S2)
    requirements = [
        ("REQ-A81-001", "Complete endpoint inventory maintained"),
        ("REQ-A81-002", "Endpoints classified by type, ownership, criticality"),
        ("REQ-A81-003", "Security baselines defined per endpoint type"),
        ("REQ-A81-004", "Baseline compliance ≥95% for corporate endpoints"),
        ("REQ-A81-005", "Full disk encryption for laptops (≥95%)"),
        ("REQ-A81-006", "Full disk encryption for desktops (≥90%)"),
        ("REQ-A81-007", "Mobile devices encrypted (≥95%)"),
        ("REQ-A81-008", "Encryption key escrow/recovery implemented"),
        ("REQ-A81-009", "MDM enrollment ≥95% corporate endpoints"),
        ("REQ-A81-010", "MDM enrollment ≥80% BYOD (if BYOD allowed)"),
        ("REQ-A81-011", "Remote wipe capability for all corporate mobile devices"),
        ("REQ-A81-012", "Screen lock enforced (max 15 minutes)"),
        ("REQ-A81-013", "Strong password policy enforced"),
        ("REQ-A81-014", "Firewall enabled on all endpoints"),
        ("REQ-A81-015", "Anti-malware deployed (see A.8.7)"),
        ("REQ-A81-016", "Automatic security updates enabled"),
        ("REQ-A81-017", "Quarterly endpoint inventory reconciliation"),
        ("REQ-A81-018", "Lost/stolen device procedure documented"),
        ("REQ-A81-019", "Lost/stolen incidents responded to within 24 hours"),
        ("REQ-A81-020", "Remote wipe initiated within 24 hours (if device not recovered)"),
        ("REQ-A81-021", "Secure disposal procedure for all endpoints"),
        ("REQ-A81-022", "Disposal certificates retained"),
        ("REQ-A81-023", "BYOD user agreements signed (if BYOD allowed)"),
        ("REQ-A81-024", "BYOD devices containerized (corporate data separated)"),
        ("REQ-A81-025", "Unmanaged endpoints identified quarterly"),
        ("REQ-A81-026", "Endpoint lifecycle management process"),
        ("REQ-A81-027", "Endpoint procurement includes security requirements"),
        ("REQ-A81-028", "Endpoints configured before user deployment"),
        ("REQ-A81-029", "Endpoint decommissioning includes data sanitization"),
        ("REQ-A81-030", "Annual endpoint security assessment"),
    ]

    start_row = 5
    for i, (req_id, requirement) in enumerate(requirements):
        current_row = start_row + i
        
        # Req ID
        ws.cell(row=current_row, column=1).value = req_id
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        # Policy Requirement
        ws.cell(row=current_row, column=2).value = requirement
        ws.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
        
        # Implemented (dropdown)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        validations['yes_no_na'].add(cell)
        
        # Evidence Reference (input)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        
        # Status (calculated)
        cell = ws.cell(row=current_row, column=6)
        cell.value = f'=IF(C{current_row}="Yes",f"{CHECK} Compliant",IF(C{current_row}="N/A","N/A",f"{XMARK} Gap"))'
        
        thin = Side(style="thin")
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Summary
    summary_row = start_row + len(requirements) + 2
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} REQUIREMENTS COMPLIANCE SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Total Requirements:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = len(requirements)
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Implemented:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(C5:C{start_row+len(requirements)-1},"Yes")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Not Implemented:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(C5:C{start_row+len(requirements)-1},"No")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Compliance Rate:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=IF(COUNT(B{summary_row-2}:B{summary_row-1})>0,B{summary_row-2}/B{summary_row-3}*100,0)&"%"'

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15


# ============================================================================
# SECTION 10: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register_sheet(ws, styles):
    """Create comprehensive evidence documentation sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "EVIDENCE REGISTER"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "Comprehensive evidence documentation for endpoint inventory assessment (100 evidence entries)"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Evidence ID",
        "Evidence Type",
        "Description",
        "Related Requirement",
        "Related Worksheet/Device",
        "File Location",
        "Collection Date",
        "Collected By",
        "Verification Status",
        "Notes"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Evidence entries (100 rows)
    start_row = 5
    for i in range(100):
        current_row = start_row + i
        
        # Evidence ID (auto-generated)
        ws.cell(row=current_row, column=1).value = f"EVD-{i+1:03d}"
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        # Evidence Type (dropdown)
        cell = ws.cell(row=current_row, column=2)
        apply_style(cell, styles['input_cell'])
        validations['evidence_type'].add(cell)
        
        # Description (input)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        
        # Related Requirement (input)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        
        # Related Worksheet/Device (input)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        
        # File Location (input)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        
        # Collection Date (input)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        
        # Collected By (input)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])
        
        # Verification Status (dropdown)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])
        validations['verification_status'].add(cell)
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])

    # Summary
    summary_row = start_row + 102
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} EVIDENCE SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Total Evidence Entries:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTA(C5:C104)'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = f"{CHECK} Verified:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(I5:I104,f"{CHECK} Verified")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "⏳ Pending:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(I5:I104,"⏳ Pending")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Verification Rate:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=IF(B{summary_row-3}>0,B{summary_row-2}/B{summary_row-3}*100,0)&"%"'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 30


# ============================================================================
# SECTION 11: GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis_sheet(ws, styles):
    """Create gap identification and remediation tracking sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:M1')
    cell = ws['A1']
    cell.value = "GAP ANALYSIS & REMEDIATION TRACKING"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:M2')
    cell = ws['A2']
    cell.value = "Gap identification, severity classification, remediation planning and tracking (40 gap entries)"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Gap ID",
        "Gap Description",
        "Affected Devices/Count",
        "Related Requirement",
        "Severity",
        "Risk",
        "Root Cause",
        "Remediation Plan",
        "Owner",
        "Due Date",
        "Status",
        "Budget Required",
        "Notes"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Gap entries (40 rows)
    start_row = 5
    for i in range(40):
        current_row = start_row + i
        
        # Gap ID (auto-generated)
        ws.cell(row=current_row, column=1).value = f"GAP-{i+1:03d}"
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        # Gap Description (input)
        cell = ws.cell(row=current_row, column=2)
        apply_style(cell, styles['input_cell'])
        
        # Affected Devices/Count (input)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        
        # Related Requirement (input)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        
        # Severity (dropdown)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        validations['gap_severity'].add(cell)
        
        # Risk (input)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        
        # Root Cause (input)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        
        # Remediation Plan (input)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])
        
        # Owner (input)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])
        
        # Due Date (input)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])
        
        # Status (dropdown)
        cell = ws.cell(row=current_row, column=11)
        apply_style(cell, styles['input_cell'])
        validations['gap_status'].add(cell)
        
        # Budget Required (input)
        cell = ws.cell(row=current_row, column=12)
        apply_style(cell, styles['input_cell'])
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=13)
        apply_style(cell, styles['input_cell'])

    # Summary
    summary_row = start_row + 42
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} GAP SUMMARY BY SEVERITY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "🔴 Critical:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E5:E44,"🔴 Critical")'
    apply_style(ws.cell(row=summary_row, column=1), styles['gap_critical'])
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "🟠 High:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E5:E44,"🟠 High")'
    apply_style(ws.cell(row=summary_row, column=1), styles['gap_high'])
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "🟡 Medium:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E5:E44,"🟡 Medium")'
    apply_style(ws.cell(row=summary_row, column=1), styles['gap_medium'])
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "🟢 Low:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(E5:E44,"🟢 Low")'
    apply_style(ws.cell(row=summary_row, column=1), styles['gap_low'])
    
    summary_row += 2
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} GAP SUMMARY BY STATUS"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "🔴 Open:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K5:K44,"🔴 Open")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "🟡 In Progress:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K5:K44,"🟡 In Progress")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = f"{CHECK} Closed:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(K5:K44,f"{CHECK} Closed")'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 35
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 30


# ============================================================================
# SECTION 12: APPROVAL SIGN-OFF SHEET
# ============================================================================

def create_approval_signoff_sheet(ws, styles):
    """Create multi-level approval workflow sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "APPROVAL & SIGN-OFF WORKFLOW"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "Three-level approval: Assessor → IT Operations Manager → CISO"
    apply_style(cell, styles['subheader'])

    # Assessment summary
    row = 4
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "📋 ASSESSMENT SUMMARY"
    apply_style(cell, styles['subheader'])
    row += 1

    summary_items = [
        ("Assessment Period:", "[Start Date] to [End Date]"),
        ("Total Endpoints Assessed:", "=COUNTA(Inventory!A5:A54)"),
        ("Baseline Compliance Rate:", "=Baseline_Compliance!B58"),
        ("Encryption Coverage Rate:", "=Encryption_Status!B58"),
        ("MDM Enrollment Rate:", "=Management_Enrollment!B58"),
        ("Requirements Implemented:", "=Capability_Requirements!B37"),
        ("Critical Gaps:", "=Gap_Analysis!B48"),
        ("High Gaps:", "=Gap_Analysis!B49"),
        ("Evidence Entries:", "=Evidence_Register!B107"),
    ]

    for label, value in summary_items:
        ws[f'A{row}'].value = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].value = value if isinstance(value, str) and value.startswith("=") else value
        row += 1

    # Approval levels
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "✍️ APPROVAL WORKFLOW"
    apply_style(cell, styles['subheader'])
    row += 1

    # Level 1: Assessor
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "LEVEL 1: ASSESSOR SIGN-OFF"
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    row += 1

    ws[f'A{row}'].value = "Assessor Name:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    row += 1

    ws[f'A{row}'].value = "Assessor Title:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    row += 1

    ws[f'A{row}'].value = "Assessment Complete Date:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    row += 1

    ws[f'A{row}'].value = "Assessor Signature:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    row += 1

    ws[f'A{row}'].value = "Assessor Comments:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    ws.merge_cells(f'B{row}:F{row}')
    ws.row_dimensions[row].height = 40
    row += 2

    # Level 2: IT Operations Manager
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "LEVEL 2: IT OPERATIONS MANAGER REVIEW"
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    row += 1

    ws[f'A{row}'].value = "Reviewer Name:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    row += 1

    ws[f'A{row}'].value = "Review Date:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    row += 1

    ws[f'A{row}'].value = "Decision:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    validations['approval_decision'].add(cell)
    row += 1

    ws[f'A{row}'].value = "IT Ops Manager Signature:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    row += 1

    ws[f'A{row}'].value = "IT Ops Manager Comments:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    ws.merge_cells(f'B{row}:F{row}')
    ws.row_dimensions[row].height = 40
    row += 2

    # Level 3: CISO
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "LEVEL 3: CISO FINAL APPROVAL"
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    row += 1

    ws[f'A{row}'].value = "CISO Name:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    row += 1

    ws[f'A{row}'].value = "Final Approval Date:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    row += 1

    ws[f'A{row}'].value = "Decision:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    validations['approval_decision'].add(cell)
    row += 1

    ws[f'A{row}'].value = "CISO Signature:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    row += 1

    ws[f'A{row}'].value = "CISO Comments:"
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    ws.merge_cells(f'B{row}:F{row}')
    ws.row_dimensions[row].height = 40
    row += 2

    # Final status
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = f"{CHART} FINAL ASSESSMENT STATUS"
    apply_style(cell, styles['subheader'])
    row += 1

    ws[f'A{row}'].value = "Assessment Status:"
    ws[f'A{row}'].font = Font(bold=True)
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])
    assessment_status_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Approved,✅ Approved with Conditions,❌ Rejected,⏳ Pending Review"',
        allow_blank=False
    )
    ws.add_data_validation(assessment_status_dv)
    assessment_status_dv.add(cell)
    
    row += 1
    ws[f'A{row}'].value = "Next Assessment Due:"
    ws[f'A{row}'].font = Font(bold=True)
    cell = ws[f'B{row}']
    apply_style(cell, styles['input_cell'])

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20


# ============================================================================
# SECTION 13: MAIN FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Systems Engineering Principle: "Don't fool yourself - and you are the 
    easiest person to fool." - Richard Feynman
    
    This script generates EVIDENCE-BASED endpoint security assessment,
    not cargo cult compliance checkbox theater.
    """
    print("=" * 78)
    print("ISMS-IMP-A.8.1-7-18-19.S1 - Endpoint Inventory Assessment Generator")
    print("ISO/IEC 27001:2022 Controls: A.8.1, A.8.7, A.8.18, A.8.19")
    print("=" * 78)
    print("\n🎯 Systems Engineering Approach: Evidence-Based Compliance")
    print(f"{CHART} Technology-Agnostic: Works with ANY endpoint environment")
    print(f"{LOCK} Audit-Ready: Comprehensive evidence collection")
    print("\n" + "─" * 78)

    # Create workbook and setup styles
    print("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    print(f"{CHECK} Workbook created with 10 sheets")

    # Create all sheets
    print("\n[Phase 2] Generating assessment sheets...")
    
    print("  [1/10] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    print("  ✅ Instructions complete")

    print("  [2/10] Creating Inventory sheet...")
    create_inventory_sheet(wb["Inventory"], styles)
    print("  ✅ Inventory complete (50 device rows)")

    print("  [3/10] Creating Classification sheet...")
    create_classification_sheet(wb["Classification"], styles)
    print("  ✅ Classification analysis complete")

    print("  [4/10] Creating Baseline_Compliance sheet...")
    create_baseline_compliance_sheet(wb["Baseline_Compliance"], styles)
    print("  ✅ Baseline compliance assessment complete")

    print("  [5/10] Creating Encryption_Status sheet...")
    create_encryption_status_sheet(wb["Encryption_Status"], styles)
    print("  ✅ Encryption status tracking complete")

    print("  [6/10] Creating Management_Enrollment sheet...")
    create_management_enrollment_sheet(wb["Management_Enrollment"], styles)
    print("  ✅ MDM enrollment tracking complete")

    print("  [7/10] Creating Capability_Requirements sheet...")
    create_capability_requirements_sheet(wb["Capability_Requirements"], styles)
    print("  ✅ Requirements mapping complete (30 A.8.1 requirements)")

    print("  [8/10] Creating Evidence_Register sheet...")
    create_evidence_register_sheet(wb["Evidence_Register"], styles)
    print("  ✅ Evidence register complete (100 evidence rows)")

    print("  [9/10] Creating Gap_Analysis sheet...")
    create_gap_analysis_sheet(wb["Gap_Analysis"], styles)
    print("  ✅ Gap analysis complete (40 gap tracking rows)")

    print("  [10/10] Creating Approval_Sign_Off sheet...")
    create_approval_signoff_sheet(wb["Approval_Sign_Of"], styles)
    print("  ✅ Approval workflow complete")

    # Save workbook
    print("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.1-7-18-19.S1_Endpoint_Inventory_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        wb.save(filename)
        print(f"{CHECK} SUCCESS: {filename}")
    except Exception as e:
        print(f"{XMARK} ERROR saving workbook: {e}")
        return 1

    # Summary
    print("\n" + "=" * 78)
    print("📋 WORKBOOK STRUCTURE SUMMARY")
    print("=" * 78)
    print("\n📄 Assessment Sheets:")
    print("  • Instructions & Legend (usage guidance, status legend)")
    print("  • Inventory (50 device rows with full metadata)")
    print("  • Classification (device type/ownership/criticality distribution)")
    print("  • Baseline_Compliance (per-endpoint baseline assessment)")
    print("  • Encryption_Status (encryption technology and key management)")
    print("  • Management_Enrollment (MDM platform and enrollment status)")
    print("\n🔧 Governance & Evidence:")
    print("  • Capability_Requirements (30 A.8.1 policy requirements)")
    print("  • Evidence_Register (100 evidence entries)")
    print("  • Gap_Analysis (40 gap tracking rows with remediation)")
    print("  • Approval_Sign_Off (3-level approval workflow)")
    print("\n" + "─" * 78)
    print("📈 ASSESSMENT CAPABILITIES:")
    print("  • 50 endpoint inventory rows (expandable)")
    print("  • 30 policy requirements mapped to implementation")
    print("  • 100 evidence documentation entries")
    print("  • 40 gap identification/remediation tracking rows")
    print("  • Automated compliance calculations")
    print("  • Multi-level approval workflow")
    print("\n" + "─" * 78)
    print(f"{TARGET} KEY FEATURES:")
    print("  ✅ Technology-agnostic (works with ANY MDM/endpoint platform)")
    print("  ✅ Comprehensive evidence collection framework")
    print("  ✅ Automated compliance scoring (baseline, encryption, enrollment)")
    print("  ✅ Gap severity classification (Critical/High/Medium/Low)")
    print("  ✅ 60+ data validations with emoji status indicators")
    print("  ✅ Audit-ready documentation")
    print("\n" + "=" * 78)
    print("🚀 NEXT STEPS:")
    print("  1. Open the generated workbook")
    print("  2. Review Instructions & Legend sheet for usage guidance")
    print("  3. Export endpoint inventory from MDM/SCCM/Jamf")
    print("  4. Populate Inventory sheet with endpoint data")
    print("  5. Complete Baseline_Compliance, Encryption_Status assessments")
    print("  6. Document evidence in Evidence_Register (min 20 entries)")
    print("  7. Identify gaps in Gap_Analysis sheet")
    print("  8. Obtain approvals via Approval_Sign_Off workflow")
    print("\n💡 PRO TIP:")
    print("  This workbook is VENDOR-AGNOSTIC. Whether you use Intune, Jamf,")
    print("  SCCM, or ANY other endpoint management platform - this framework")
    print("  assesses CAPABILITIES and COMPLIANCE, not brand names.")
    print("\n" + "=" * 78)
    print('\n"The first principle is that you must not fool yourself')
    print('— and you are the easiest person to fool." - Richard Feynman')
    print("\n🎓 This is not cargo cult ISMS. This is evidence-based compliance.")
    print("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    exit(main())
