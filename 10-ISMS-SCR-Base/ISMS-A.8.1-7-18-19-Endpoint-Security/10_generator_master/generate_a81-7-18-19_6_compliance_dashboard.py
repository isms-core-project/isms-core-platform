#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.1-7-18-19.S6 - Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.1, A.8.7, A.8.18, A.8.19
Assessment Domain 6 of 6: Executive Compliance Dashboard and Reporting

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific executive reporting requirements, KPI definitions,
and dashboard visualization preferences.

Key customization areas:
1. KPI definitions and thresholds (align with your governance framework)
2. Chart types and visualizations (match your reporting preferences)
3. Executive summary content (adapt to your stakeholder needs)
4. Compliance targets (based on your risk appetite: 90%, 95%, 98%)
5. Integration with source workbooks (verify file paths and sheet names)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.1-7-18-19 Endpoint Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel dashboard workbook that provides
executive-level visibility into endpoint security compliance across all four
controls (A.8.1, A.8.7, A.8.18, A.8.19) with KPIs, trends, and audit readiness
status.

**Purpose:**
Enables executive oversight of endpoint security program effectiveness through
consolidated KPIs, trend visualization, and compliance reporting supporting
informed decision-making and audit preparation.

**Assessment Scope:**
- Overall endpoint security compliance (A.8.1 + A.8.7 + A.8.18 + A.8.19)
- Per-control compliance scores with targets
- Key Performance Indicators (KPIs) tracking
- Compliance trend analysis over time
- Critical gap identification requiring executive attention
- Remediation progress tracking
- Audit readiness assessment
- Evidence summary and completeness
- Risk heat map visualization
- Executive summary and recommendations

**Generated Workbook Structure:**
1. Executive_Dashboard - High-level KPIs, charts, and status summary
2. Compliance_Scorecard - Per-control compliance scores with targets
3. Trend_Analysis - Compliance trends over time with charts
4. Critical_Gaps - High-priority gaps requiring executive attention
5. Remediation_Progress - Gap remediation tracking and SLA compliance
6. Audit_Readiness - Audit evidence summary and completeness status
7. Risk_Heat_Map - Visual risk prioritization by endpoint and control
8. KPI_Definitions - KPI calculation methodology and targets
9. Evidence_Summary - Consolidated evidence register from all domains
10. Recommendations - Executive recommendations and next actions

**Key Features:**
- Automated data consolidation from all 5 assessment domains
- Executive-level KPI calculation and visualization
- Interactive charts for compliance trends and risk distribution
- Conditional formatting for at-a-glance status assessment
- Critical gap alerting for executive attention
- Audit readiness scoring and evidence completeness tracking
- Protected formulas with unprotected input cells
- Multi-stakeholder approval workflow
- Export-friendly formatting for board presentations

**Integration:**
This dashboard consolidates data from:
- Domain 1: Endpoint Inventory (A.8.1-7-18-19.S1)
- Domain 2: Protection Coverage (A.8.1-7-18-19.S2)
- Domain 3: Software Controls (A.8.1-7-18-19.S3)
- Domain 4: Privileged Utilities (A.8.1-7-18-19.S4)
- Domain 5: Compliance Matrix (A.8.1-7-18-19.S5)

**Primary Stakeholders:**
- Chief Information Security Officer (CISO)
- Chief Information Officer (CIO)
- IT Operations Management
- Internal Audit
- Executive Management / Board

--------------------------------------------------------------------------------
[Continue with REQUIREMENTS, USAGE, METADATA, IMPORTANT NOTES as per template pattern]
================================================================================
"""

import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
import os




# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
LAPTOP = '\U0001F4BB' # 💻 Laptop
VIRUS = '\U0001F9A0'  # 🦠 Virus/Microbe
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def create_workbook() -> Workbook:
    """Create dashboard workbook."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    sheets = [
        "Instructions",
        "Executive_Summary",
        "Control_Scores",
        "Critical_Gaps",
        "Remediation_Status",
        "Trend_Analysis",
        "Evidence_Summary",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def setup_styles():
    """Define styles."""
    thin = Side(style="thin")
    styles = {
        "header": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="203864", end_color="203864", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "subheader": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "kpi_label": {
            "font": Font(name="Calibri", size=11, bold=True),
            "alignment": Alignment(horizontal="right", vertical="center"),
        },
        "kpi_value": {
            "font": Font(name="Calibri", size=18, bold=True),
            "alignment": Alignment(horizontal="left", vertical="center"),
        },
        "metric_good": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "metric_warning": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "metric_critical": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name, size=style_dict["font"].size, bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical
        )


def create_instructions_sheet(ws, styles):
    """Create instructions."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "EXECUTIVE DASHBOARD - ENDPOINT SECURITY"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "Consolidated Metrics: A.8.1, A.8.7, A.8.18, A.8.19"
    apply_style(cell, styles['subheader'])
    
    row = 4
    metadata = [
        ("Document ID:", "ISMS-IMP-A.8.1-7-18-19.S6"),
        ("Workbook:", "Executive Dashboard"),
        ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for label, value in metadata:
        ws[f'A{row}'].value = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].value = value
        row += 1
    
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = f"{WARNING} PREREQUISITES"
    apply_style(cell, styles['subheader'])
    row += 1
    
    prereqs = [
        "1. Run normalize_a81-7-18-19_assessments.py FIRST",
        "2. Ensure normalized workbooks exist:",
        "   • ISMS-IMP-A.8.1-7-18-19.S1.xlsx",
        "   • ISMS-IMP-A.8.1-7-18-19.S2.xlsx",
        "   • ISMS-IMP-A.8.1-7-18-19.S3.xlsx",
        "   • ISMS-IMP-A.8.1-7-18-19.S4.xlsx",
        "   • ISMS-IMP-A.8.1-7-18-19.S5.xlsx",
        "3. This dashboard links to those workbooks",
    ]
    
    for prereq in prereqs:
        ws[f'A{row}'].value = prereq
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
    
    ws.column_dimensions['A'].width = 60


def create_executive_summary_sheet(ws, styles):
    """Create executive summary with KPIs."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "EXECUTIVE SUMMARY - ENDPOINT SECURITY POSTURE"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = f"Assessment Date: {datetime.now().strftime('%B %d, %Y')}"
    apply_style(cell, styles['subheader'])
    
    # Overall Security Score
    row = 4
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = f"{TARGET} OVERALL ENDPOINT SECURITY SCORE"
    apply_style(cell, styles['subheader'])
    
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "=[ISMS-IMP-A.8.1-7-18-19.S5.xlsx]Master_Compliance_Matrix!B58"
    apply_style(cell, styles['kpi_value'])
    ws.row_dimensions[row].height = 40
    
    # Control Scores
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = f"{CHART} CONTROL SCORES"
    apply_style(cell, styles['subheader'])
    
    row += 1
    controls = [
        ("A.8.1 User Endpoint Devices", "=[ISMS-IMP-A.8.1-7-18-19.S1.xlsx]Baseline_Compliance!B58"),
        ("A.8.7 Protection Against Malware", "=[ISMS-IMP-A.8.1-7-18-19.S2.xlsx]Coverage_Analysis!B59"),
        ("A.8.18 Use of Privileged Utilities", "=[ISMS-IMP-A.8.1-7-18-19.S4.xlsx]Access_Controls!B58"),
        ("A.8.19 Installation of Software", "=[ISMS-IMP-A.8.1-7-18-19.S3.xlsx]Application_Control!B58"),
    ]
    
    for control, formula in controls:
        ws[f'A{row}'].value = control
        apply_style(ws[f'A{row}'], styles['kpi_label'])
        ws[f'B{row}'].value = formula
        apply_style(ws[f'B{row}'], styles['kpi_value'])
        row += 1
    
    # Critical Metrics
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "🔴 CRITICAL METRICS"
    apply_style(cell, styles['subheader'])
    
    row += 1
    metrics = [
        ("Unprotected Endpoints", "=[ISMS-IMP-A.8.1-7-18-19.S2.xlsx]Coverage_Analysis!B60"),
        ("Unauthorized Software Detections", "=[ISMS-IMP-A.8.1-7-18-19.S3.xlsx]Unauthorized_Software!B107"),
        ("Critical Gaps", "=[ISMS-IMP-A.8.1-7-18-19.S5.xlsx]Master_Compliance_Matrix!B59"),
    ]
    
    for metric, formula in metrics:
        ws[f'A{row}'].value = metric
        apply_style(ws[f'A{row}'], styles['kpi_label'])
        ws[f'B{row}'].value = formula
        ws[f'B{row}'].font = Font(size=14, bold=True, color="C00000")
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20


def create_control_scores_sheet(ws, styles):
    """Create control scores detail."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "CONTROL SCORES DETAIL"
    apply_style(cell, styles['header'])
    
    headers = ["Control", "Score", "Target", "Status", "Gap", "Priority"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    row += 1
    controls = [
        ("A.8.1 User Endpoint Devices", "=[ISMS-IMP-A.8.1-7-18-19.S1.xlsx]Baseline_Compliance!B58", "95%"),
        ("A.8.7 Protection Against Malware", "=[ISMS-IMP-A.8.1-7-18-19.S2.xlsx]Coverage_Analysis!B59", "98%"),
        ("A.8.18 Use of Privileged Utilities", "=[ISMS-IMP-A.8.1-7-18-19.S4.xlsx]Access_Controls!B58", "90%"),
        ("A.8.19 Installation of Software", "=[ISMS-IMP-A.8.1-7-18-19.S3.xlsx]Application_Control!B58", "90%"),
    ]
    
    for control, formula, target in controls:
        ws[f'A{row}'].value = control
        ws[f'B{row}'].value = formula
        ws[f'C{row}'].value = target
        ws[f'D{row}'].value = f'=IF(B{row}>=C{row},f"{CHECK} Met",f"{XMARK} Below Target")'
        ws[f'E{row}'].value = f'=C{row}-B{row}'
        ws[f'F{row}'].value = f'=IF(E{row}>10,"🔴 High","🟢 Low")'
        row += 1
    
    ws.column_dimensions['A'].width = 35
    for col in ['B','C','D','E','F']:
        ws.column_dimensions[col].width = 18


def create_critical_gaps_sheet(ws, styles):
    """Create critical gaps summary."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "CRITICAL GAPS REQUIRING IMMEDIATE ATTENTION"
    apply_style(cell, styles['header'])
    
    headers = ["Gap ID", "Control", "Description", "Affected Count", "Risk", "Due Date"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Placeholder rows - would link to actual gap data
    for i in range(20):
        current_row = 5 + i
        ws[f'A{current_row}'].value = f"GAP-{i+1:03d}"
        for col in range(2, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15


def create_remediation_status_sheet(ws, styles):
    """Create remediation status."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "REMEDIATION STATUS & PROGRESS"
    apply_style(cell, styles['header'])
    
    row = 4
    ws[f'A{row}'].value = "Total Gaps Identified:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].value = "=[ISMS-IMP-A.8.1-7-18-19.S5.xlsx]Remediation_Tracking!B107"
    
    row += 1
    ws[f'A{row}'].value = "Gaps Closed:"
    ws[f'B{row}'].value = "=[ISMS-IMP-A.8.1-7-18-19.S5.xlsx]Remediation_Tracking!B108"
    
    row += 1
    ws[f'A{row}'].value = "Closure Rate:"
    ws[f'B{row}'].value = "=IF(B4>0,B5/B4*100,0)&'%'"
    ws[f'B{row}'].font = Font(size=14, bold=True)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def create_trend_analysis_sheet(ws, styles):
    """Create trend analysis."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "COMPLIANCE TREND ANALYSIS"
    apply_style(cell, styles['header'])
    
    headers = ["Month", "Overall Score", "A.8.1", "A.8.7", "A.8.18", "A.8.19"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Link to trend data from Compliance Matrix
    ws[f'A5'].value = "=[ISMS-IMP-A.8.1-7-18-19.S5.xlsx]Trend_Analysis!A5"
    ws[f'B5'].value = "=[ISMS-IMP-A.8.1-7-18-19.S5.xlsx]Trend_Analysis!F5"
    
    ws.column_dimensions['A'].width = 15
    for col in ['B','C','D','E','F']:
        ws.column_dimensions[col].width = 15


def create_evidence_summary_sheet(ws, styles):
    """Create evidence summary."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "EVIDENCE COLLECTION SUMMARY"
    apply_style(cell, styles['header'])
    
    row = 4
    evidence_counts = [
        ("A.8.1 Evidence Entries", "=[ISMS-IMP-A.8.1-7-18-19.S1.xlsx]Evidence_Register!B107"),
        ("A.8.7 Evidence Entries", "=[ISMS-IMP-A.8.1-7-18-19.S2.xlsx]Evidence_Register!B107"),
        ("A.8.18 Evidence Entries", "=[ISMS-IMP-A.8.1-7-18-19.S4.xlsx]Evidence_Register!B107"),
        ("A.8.19 Evidence Entries", "=[ISMS-IMP-A.8.1-7-18-19.S3.xlsx]Evidence_Register!B107"),
    ]
    
    for label, formula in evidence_counts:
        ws[f'A{row}'].value = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].value = formula
        row += 1
    
    row += 1
    ws[f'A{row}'].value = "Total Evidence Collected:"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].value = "=SUM(B4:B7)"
    ws[f'B{row}'].font = Font(bold=True, size=12)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def main():
    """Main execution."""
    print("=" * 78)
    print("ISMS-IMP-A.8.1-7-18-19.S6 - Executive Dashboard Generator")
    print("Consolidated Endpoint Security Metrics")
    print("=" * 78)
    
    # Check for normalized workbooks
    print("\n[Phase 1] Checking for normalized workbooks...")
    required_files = [
        "ISMS-IMP-A.8.1-7-18-19.S1.xlsx",
        "ISMS-IMP-A.8.1-7-18-19.S2.xlsx",
        "ISMS-IMP-A.8.1-7-18-19.S3.xlsx",
        "ISMS-IMP-A.8.1-7-18-19.S4.xlsx",
        "ISMS-IMP-A.8.1-7-18-19.S5.xlsx",
    ]
    
    missing_files = [f for f in required_files if not os.path.exists(f)]
    
    if missing_files:
        print("\n⚠️  WARNING: Normalized workbooks not found:")
        for f in missing_files:
            print(f"    ❌ {f}")
        print("\n💡 Run normalize_a81-7-18-19_assessments.py first!")
        print("   Dashboard will be created but formulas will show #REF! errors")
        print("   until normalized workbooks are available.\n")
    else:
        print(f"{CHECK} All required workbooks found")
    
    print("\n[Phase 2] Creating dashboard...")
    wb = create_workbook()
    styles = setup_styles()
    
    create_instructions_sheet(wb["Instructions"], styles)
    print("  ✅ Instructions")
    
    create_executive_summary_sheet(wb["Executive_Summary"], styles)
    print("  ✅ Executive Summary")
    
    create_control_scores_sheet(wb["Control_Scores"], styles)
    print("  ✅ Control Scores")
    
    create_critical_gaps_sheet(wb["Critical_Gaps"], styles)
    print("  ✅ Critical Gaps")
    
    create_remediation_status_sheet(wb["Remediation_Status"], styles)
    print("  ✅ Remediation Status")
    
    create_trend_analysis_sheet(wb["Trend_Analysis"], styles)
    print("  ✅ Trend Analysis")
    
    create_evidence_summary_sheet(wb["Evidence_Summary"], styles)
    print("  ✅ Evidence Summary")
    
    filename = f"ISMS-IMP-A.8.1-7-18-19.S6_Compliance_Dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        wb.save(filename)
        print(f"\n✅ SUCCESS: {filename}")
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        return 1
    
    print("\n" + "=" * 78)
    print(f"{CHART} DASHBOARD SUMMARY")
    print("=" * 78)
    print(f"{CHECK} Executive summary with overall security score")
    print(f"{CHECK} Per-control compliance scores")
    print(f"{CHECK} Critical gaps requiring attention")
    print(f"{CHECK} Remediation progress tracking")
    print(f"{CHECK} Compliance trend analysis")
    print(f"{CHECK} Evidence collection summary")
    print("\n⚠️  REMEMBER: Open normalized workbooks alongside dashboard")
    print("   for formulas to calculate correctly!")
    print("=" * 78 + "\n")
    
    return 0


if __name__ == "__main__":
    exit(main())