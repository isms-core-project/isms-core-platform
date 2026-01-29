#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.1-7-18-19.S4 - Privileged Utilities Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.1, A.8.7, A.8.18, A.8.19
Assessment Domain 4 of 6: Privileged Utility Programs and Administrative Tools

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific privileged utility landscape, access control
mechanisms, and monitoring requirements.

Key customization areas:
1. Privileged utility inventory (match your actual administrative tools)
2. Access control mechanisms (PAM, role-based access, etc.)
3. Usage monitoring and logging (specific to your SIEM integration)
4. Approval workflows (aligned with your privileged access governance)
5. Just-in-time access implementation (if applicable)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.1-7-18-19 Endpoint Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
privileged utility program controls, access management, and usage monitoring
across all endpoint devices.

**Purpose:**
Enables systematic assessment of privileged utility governance and monitoring
against ISO 27001:2022 Control A.8.18 requirements, supporting evidence-based
validation of privileged access controls.

**Assessment Scope:**
- Privileged utility inventory (system tools, admin utilities, debugging tools)
- Access control mechanisms for privileged utilities
- Usage monitoring and audit logging
- Approval workflow compliance for privileged access
- Tools that can bypass security controls (identification and restriction)
- Administrative tool management (remote access, system utilities)
- Debugging and development tools on production endpoints
- Just-in-time privileged access (if applicable)
- Gap analysis for uncontrolled privileged utilities
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and privileged utility framework
2. Privileged_Utility_Inventory - Master inventory of privileged tools
3. Access_Controls - Per-utility access control mechanisms
4. Usage_Monitoring - Privileged utility usage logging and monitoring
5. Approval_Compliance - Approval workflow tracking
6. Bypass_Tool_Analysis - Tools that can bypass security controls
7. Administrative_Tools - Remote access and system utility management
8. Evidence_Register - Audit evidence tracking and documentation
9. Gap_Analysis - Uncontrolled or unmonitored privileged utilities
10. Remediation_Tracking - Privileged utility control gap remediation
11. Approval_Sign_Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with privileged utility category dropdown lists
- Conditional formatting for access control compliance visualization
- Automated gap identification for unmonitored utilities
- Usage audit trail tracking
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with SIEM and PAM solutions

**Integration:**
This assessment builds on the endpoint inventory (A.8.1-7-18-19.S1) and feeds
into the consolidated compliance matrix (A.8.1-7-18-19.S5) and dashboard
(A.8.1-7-18-19.S6).

--------------------------------------------------------------------------------
[Continue with REQUIREMENTS, USAGE, METADATA, IMPORTANT NOTES as per template pattern]
================================================================================
"""

import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation




# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
LAPTOP = '\U0001F4BB' # 💻 Laptop
VIRUS = '\U0001F9A0'  # 🦠 Virus/Microbe
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    sheets = [
        "Instructions & Legend",
        "Utility_Inventory",
        "Access_Controls",
        "Approval_Workflow",
        "Usage_Audit",
        "MFA_Compliance",
        "Quarterly_Reviews",
        "Capability_Requirements",
        "Evidence_Register",
        "Gap_Analysis",
        "Approval_Sign_Of",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def setup_styles():
    """Define cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="44546A", end_color="44546A", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "status_controlled": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        },
        "status_uncontrolled": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "gap_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "gap_high": {"fill": PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")},
        "gap_medium": {"fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")},
        "gap_low": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style to cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name, size=style_dict["font"].size, bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def create_base_validations(ws):
    """Create data validations."""
    validations = {
        'yes_no': DataValidation(type="list", formula1='"Yes,No"', allow_blank=False),
        'yes_no_na': DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=False),
        'utility_category': DataValidation(type="list", formula1='"🛠️ System Admin,🐛 Debugging,🔓 Security Bypass,🌐 Network Tools,💾 Disk/File,🔧 Third-Party Admin,Other"', allow_blank=False),
        'risk_level': DataValidation(type="list", formula1='"🔴 Critical,🟠 High,🟡 Medium,🟢 Low"', allow_blank=False),
        'access_control_status': DataValidation(type="list", formula1=f'"{CHECK} Controlled,⚠️ Partial,❌ Uncontrolled,❓ Unknown"', allow_blank=False),
        'access_type': DataValidation(type="list", formula1='"Standing,Temporary,JIT High-Risk,JIT Critical,Emergency"', allow_blank=False),
        'approval_status': DataValidation(type="list", formula1=f'"{CHECK} Approved,⏳ Pending,❌ Rejected"', allow_blank=False),
        'mfa_status': DataValidation(type="list", formula1=f'"{CHECK} Enabled,❌ Not Enabled,N/A"', allow_blank=False),
        'review_decision': DataValidation(type="list", formula1=f'"{CHECK} Continue,❌ Revoke,🔄 Modify"', allow_blank=False),
        'gap_severity': DataValidation(type="list", formula1='"🔴 Critical,🟠 High,🟡 Medium,🟢 Low"', allow_blank=False),
        'gap_status': DataValidation(type="list", formula1='"🔴 Open,🟡 In Progress,🟢 Resolved,✅ Closed"', allow_blank=False),
        'evidence_type': DataValidation(type="list", formula1='"📄 Config,📸 Screenshot,📊 Report,📋 Policy,📁 Log,Other"', allow_blank=False),
        'verification_status': DataValidation(type="list", formula1=f'"{CHECK} Verified,⏳ Pending,❌ Not Verified"', allow_blank=False),
        'approval_decision': DataValidation(type="list", formula1=f'"{CHECK} Approved,✅ Approved with Conditions,❌ Rejected,⏳ Pending"', allow_blank=False),
    }
    for dv in validations.values():
        ws.add_data_validation(dv)
    return validations


def create_instructions_sheet(ws, styles):
    """Create instructions sheet."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "PRIVILEGED UTILITY MANAGEMENT & ACCESS CONTROL"
    apply_style(cell, styles['header'])
    
    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "ISO/IEC 27001:2022 - Control A.8.18 (Use of Privileged Utility Programs)"
    apply_style(cell, styles['subheader'])
    
    row = 4
    metadata = [
        ("Document ID:", "ISMS-IMP-A.8.1-7-18-19.S4"),
        ("Workbook:", "Privileged Utilities Assessment"),
        ("Version:", "1.0"),
        ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for label, value in metadata:
        ws[f'A{row}'].value = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].value = value
        row += 1
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60


def create_utility_inventory_sheet(ws, styles):
    """Create privileged utility inventory."""
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "PRIVILEGED UTILITY INVENTORY"
    apply_style(cell, styles['header'])
    
    headers = ["Utility ID", "Utility Name", "Platform", "Category", "Risk Level", "Business Justification", "Authorized Roles", "Access Control Method", "MFA Required", "Logging Enabled", "Notes"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    for i in range(100):
        current_row = 5 + i
        ws.cell(row=current_row, column=1).value = f"UTIL-{i+1:03d}"
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        for col in range(2, 12):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
            if col == 4:
                validations['utility_category'].add(cell)
            elif col == 5:
                validations['risk_level'].add(cell)
            elif col in [9, 10]:
                validations['yes_no'].add(cell)
    
    summary_row = 107
    ws[f'A{summary_row}'].value = "Total Utilities:"
    ws[f'B{summary_row}'].value = f'=COUNTA(B5:B104)'
    ws[f'A{summary_row}'].font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 30


def create_access_controls_sheet(ws, styles):
    """Create access control configuration sheet."""
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "ACCESS CONTROL CONFIGURATION"
    apply_style(cell, styles['header'])
    
    headers = ["Utility ID", "Utility Name", "File Permissions Set", "AD Group Restriction", "AppLocker Rule", "Control Status", "Last Verified", "Verified By", "Issues", "Notes"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    for i in range(100):
        current_row = 5 + i
        ws.cell(row=current_row, column=1).value = f"UTIL-{i+1:03d}"
        
        for col in range(2, 11):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
            if col in [3, 4, 5]:
                validations['yes_no_na'].add(cell)
            elif col == 6:
                validations['access_control_status'].add(cell)
    
    ws.column_dimensions['A'].width = 12
    for col in ['B','C','D','E','F','G','H','I','J']:
        ws.column_dimensions[col].width = 20


def create_approval_workflow_sheet(ws, styles):
    """Create approval workflow tracking."""
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "PRIVILEGED ACCESS APPROVAL WORKFLOW"
    apply_style(cell, styles['header'])
    
    headers = ["Request ID", "Request Date", "Requester", "Utility", "Access Type", "Duration", "Justification", "Manager Approval", "Security Approval", "Status", "Notes"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    for i in range(100):
        current_row = 5 + i
        ws.cell(row=current_row, column=1).value = f"REQ-{i+1:04d}"
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        for col in range(2, 12):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
            if col == 5:
                validations['access_type'].add(cell)
            elif col in [8, 9]:
                validations['yes_no'].add(cell)
            elif col == 10:
                validations['approval_status'].add(cell)
    
    ws.column_dimensions['A'].width = 12
    for col in ['B','C','D','E','F','G','H','I','J','K']:
        ws.column_dimensions[col].width = 20


def create_usage_audit_sheet(ws, styles):
    """Create usage audit log sheet."""
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "PRIVILEGED UTILITY USAGE AUDIT LOG"
    apply_style(cell, styles['header'])
    
    headers = ["Log ID", "Timestamp", "User", "Utility", "Device", "Action", "Duration", "Authorized", "Flagged", "Notes"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    for i in range(200):
        current_row = 5 + i
        ws.cell(row=current_row, column=1).value = f"LOG-{i+1:05d}"
        for col in range(2, 11):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
    
    ws.column_dimensions['A'].width = 12
    for col in ['B','C','D','E','F','G','H','I','J']:
        ws.column_dimensions[col].width = 20


def create_mfa_compliance_sheet(ws, styles):
    """Create MFA compliance tracking."""
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "MFA COMPLIANCE FOR PRIVILEGED ACCESS"
    apply_style(cell, styles['header'])
    
    headers = ["User", "Role", "Utility Access", "MFA Technology", "MFA Status", "Last MFA Setup", "Compliance", "Notes"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    for i in range(50):
        current_row = 5 + i
        for col in range(1, 9):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
            if col == 5:
                validations['mfa_status'].add(cell)
    
    ws.column_dimensions['A'].width = 20
    for col in ['B','C','D','E','F','G','H']:
        ws.column_dimensions[col].width = 20


def create_quarterly_reviews_sheet(ws, styles):
    """Create quarterly access review tracking."""
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "QUARTERLY PRIVILEGED ACCESS REVIEWS"
    apply_style(cell, styles['header'])
    
    headers = ["Review Quarter", "User", "Utility Access", "Last Used", "Manager Review", "Decision", "Action Taken", "Completion Date", "Reviewer", "Notes"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    for i in range(100):
        current_row = 5 + i
        for col in range(1, 11):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
            if col == 6:
                validations['review_decision'].add(cell)
    
    ws.column_dimensions['A'].width = 15
    for col in ['B','C','D','E','F','G','H','I','J']:
        ws.column_dimensions[col].width = 20


def create_capability_requirements_sheet(ws, styles):
    """Create A.8.18 requirements mapping."""
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "CAPABILITY REQUIREMENTS MAPPING (A.8.18)"
    apply_style(cell, styles['header'])
    
    headers = ["Req ID", "Policy Requirement", "Implemented", "Evidence Reference", "Notes", "Status"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    requirements = [
        ("REQ-A818-001", "Privileged utility inventory maintained"),
        ("REQ-A818-002", "Utilities classified by risk level"),
        ("REQ-A818-003", "Access controls implemented for all privileged utilities"),
        ("REQ-A818-004", "Approval workflow for privileged access"),
        ("REQ-A818-005", "MFA required for high-risk utilities ≥90%"),
        ("REQ-A818-006", "Logging enabled for all privileged utility usage"),
        ("REQ-A818-007", "SIEM integration for privileged utility logs"),
        ("REQ-A818-008", "Quarterly access reviews conducted"),
        ("REQ-A818-009", "JIT access for critical utilities"),
        ("REQ-A818-010", "Security bypass tools restricted"),
    ]
    
    start_row = 5
    for i, (req_id, requirement) in enumerate(requirements):
        current_row = start_row + i
        ws.cell(row=current_row, column=1).value = req_id
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2).value = requirement
        ws.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
        
        for col in range(3, 7):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
            if col == 3:
                validations['yes_no_na'].add(cell)
            elif col == 6:
                cell.value = f'=IF(C{current_row}="Yes",f"{CHECK} Compliant",f"{XMARK} Gap")'
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15


def create_evidence_register_sheet(ws, styles):
    """Create evidence register."""
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "EVIDENCE REGISTER"
    apply_style(cell, styles['header'])
    
    headers = ["Evidence ID", "Type", "Description", "Related Requirement", "Worksheet", "Location", "Date", "Collected By", "Status", "Notes"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    for i in range(100):
        current_row = 5 + i
        ws.cell(row=current_row, column=1).value = f"EVD-{i+1:03d}"
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        for col in range(2, 11):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
            if col == 2:
                validations['evidence_type'].add(cell)
            elif col == 9:
                validations['verification_status'].add(cell)
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40


def create_gap_analysis_sheet(ws, styles):
    """Create gap analysis."""
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:L1')
    cell = ws['A1']
    cell.value = "GAP ANALYSIS & REMEDIATION"
    apply_style(cell, styles['header'])
    
    headers = ["Gap ID", "Description", "Affected Utilities", "Requirement", "Severity", "Risk", "Root Cause", "Remediation", "Owner", "Due Date", "Status", "Notes"]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    for i in range(40):
        current_row = 5 + i
        ws.cell(row=current_row, column=1).value = f"GAP-{i+1:03d}"
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        for col in range(2, 13):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
            if col == 5:
                validations['gap_severity'].add(cell)
            elif col == 11:
                validations['gap_status'].add(cell)
    
    ws.column_dimensions['A'].width = 12
    for col in ['B','C','D','E','F','G','H','I','J','K','L']:
        ws.column_dimensions[col].width = 20


def create_approval_signoff_sheet(ws, styles):
    """Create approval workflow."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "APPROVAL & SIGN-OFF"
    apply_style(cell, styles['header'])
    
    ws[f'A4'].value = "Assessment Summary:"
    ws[f'A4'].font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30


def main():
    """Main execution."""
    print("=" * 78)
    print("ISMS-IMP-A.8.1-7-18-19.S4 - Privileged Utilities Assessment Generator")
    print("ISO/IEC 27001:2022 Control: A.8.18")
    print("=" * 78)
    
    wb = create_workbook()
    styles = setup_styles()
    
    sheets = [
        ("Instructions & Legend", create_instructions_sheet),
        ("Utility_Inventory", create_utility_inventory_sheet),
        ("Access_Controls", create_access_controls_sheet),
        ("Approval_Workflow", create_approval_workflow_sheet),
        ("Usage_Audit", create_usage_audit_sheet),
        ("MFA_Compliance", create_mfa_compliance_sheet),
        ("Quarterly_Reviews", create_quarterly_reviews_sheet),
        ("Capability_Requirements", create_capability_requirements_sheet),
        ("Evidence_Register", create_evidence_register_sheet),
        ("Gap_Analysis", create_gap_analysis_sheet),
        ("Approval_Sign_Of", create_approval_signoff_sheet),
    ]
    
    for i, (name, func) in enumerate(sheets, 1):
        print(f"  [{i}/11] Creating {name}...")
        func(wb[name], styles)
    
    filename = f"ISMS-IMP-A.8.1-7-18-19.S4_Privileged_Utilities_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        wb.save(filename)
        print(f"{CHECK} SUCCESS: {filename}")
    except Exception as e:
        print(f"{XMARK} ERROR: {e}")
        return 1
    
    print("\n✅ 11 sheets: Inventory, Access Controls, Approvals, Usage Audit, MFA")
    print(f"{CHECK} Evidence-based privileged utility management assessment")
    return 0


if __name__ == "__main__":
    exit(main())
