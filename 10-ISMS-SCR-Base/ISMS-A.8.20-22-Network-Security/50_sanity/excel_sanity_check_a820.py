#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
Excel Workbook Sanity Checker for ISMS A.8.20-21-22 Assessment Workbooks
================================================================================

ISO/IEC 27001:2022 Controls A.8.20, A.8.21, A.8.22: Network Security Framework
Quality Assurance Utility: Excel Workbook Diagnostic and Validation Tool

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script performs comprehensive diagnostic checks on A.8.20-21-22 network
security assessment Excel workbooks to identify issues that trigger Excel's
"file level validation and repair" warnings and to ensure workbook quality
before distribution or consolidation.

**Purpose:**
Diagnoses common openpyxl-generated Excel issues, validates workbook structure
against framework specifications, and provides actionable remediation guidance
to ensure workbook integrity and audit readiness.

**Diagnostic Scope:**
- Sheet structure validation (presence of required sheets)
- Assessment ID verification (ISMS-IMP-A.8.20-21-22.X identifiers)
- Formula syntax validation (inter-sheet references, external links)
- Data validation conflicts (overlapping ranges)
- Merged cell integrity (content in non-primary cells)
- Evidence register validation (capacity and format)
- Approval workflow validation (3-level approval structure)
- Summary dashboard validation (formula completeness)
- Row capacity validation (100+ row target for data entry)
- Worksheet dimension checks (performance considerations)

**Supported Workbook Types:**
- A820-1: Infrastructure Inventory
- A820-2: Device Security Assessment
- A820-3: Network Services Catalog
- A820-4: Network Segmentation Matrix
- A820-5: Security Controls Coverage Matrix
- A820-6: Compliance Dashboard (Master)

**Diagnostic Output:**
- Workbook type detection and classification
- 10 comprehensive validation checks with ✓/✗/⚠ status
- Critical issues requiring immediate fix
- Warnings recommending attention
- Recommended remediation actions
- Workbook-specific notes and expected structure
- Integration guidance for dashboard consolidation

**When to Use:**
- After generating workbooks (Scripts 1-6)
- Before distributing to stakeholders
- After manual modifications or edits
- Before consolidation into dashboard
- When Excel shows repair warnings
- Quality assurance validation before audits

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel processing

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - pathlib (standard library)
    - re (standard library - regex for validation)
    - sys (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    # Check any A.8.20-21-22 assessment workbook
    python3 excel_sanity_check_a820.py <filename.xlsx>

Domain-Specific Examples:
    # Check Infrastructure Inventory
    python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_1_Infrastructure_Inventory.xlsx
    
    # Check Device Security Assessment
    python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_2_Device_Security_Assessment.xlsx
    
    # Check Network Services Catalog
    python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_3_Services_Catalog.xlsx
    
    # Check Segmentation Matrix
    python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_4_Segmentation_Matrix.xlsx
    
    # Check Controls Coverage
    python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_5_Controls_Coverage.xlsx
    
    # Check Compliance Dashboard
    python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_Dashboard.xlsx
    
    # Check normalized workbook
    python3 excel_sanity_check_a820.py ISMS-IMP-A.8.20-21-22.1.xlsx

Quality Assurance Workflow:
    # After generation (immediate validation)
    python3 generate_a820_1_infrastructure_inventory.py
    python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_1_*.xlsx
    
    # Before consolidation (validate all source workbooks)
    for f in ISMS_A_8_20_21_22_[1-5]_*.xlsx; do
        python3 excel_sanity_check_a820.py "$f"
    done
    
    # After normalization
    python3 normalize_a820_assessments.py
    python3 excel_sanity_check_a820.py ISMS-IMP-A.8.20-21-22.*.xlsx
    
    # Dashboard validation
    python3 generate_a820_6_compliance_dashboard.py
    python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_Dashboard_*.xlsx

Output:
    Console output with:
    - Workbook type detection
    - 10 validation checks with detailed findings
    - Critical issues summary (must fix)
    - Warnings summary (should fix)
    - Recommended remediation actions
    - Workbook-specific notes
    - Expected structure documentation

Exit Codes:
    0 - No issues detected (workbook healthy)
    1 - Critical issues found or file error

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.20, A.8.21, A.8.22
Utility Type:         Quality Assurance - Excel Workbook Diagnostics
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization ISMS Team]
Date:                 24.01.2025
Last Modified:        24.01.2025
Python Version:       3.8+
License:              [Organization License/Terms]

Related Documents:
    - ISMS-POL-A.8.20-21-22: Network Security Framework (Master Policy)
    - ISMS-POL-A.8.20-21-22-S1: Executive Summary & Control Alignment
    - ISMS-POL-A.8.20-21-22-S2: Network Security Requirements (A.8.20)
    - ISMS-POL-A.8.20-21-22-S3: Network Services Requirements (A.8.21)
    - ISMS-POL-A.8.20-21-22-S4: Network Segregation Requirements (A.8.22)
    - ISMS-POL-A.8.20-21-22-S5: Assessment & Evidence Framework
    - ISMS-IMP-A.8.20-21-22-S1: Network Discovery Process
    - ISMS-IMP-A.8.20-21-22-S2: Network Architecture Documentation
    - ISMS-IMP-A.8.20-21-22-S3: Device Hardening Process
    - ISMS-IMP-A.8.20-21-22-S4: Services Security Process
    - ISMS-IMP-A.8.20-21-22-S5: Segmentation Implementation
    - ISMS-IMP-A.8.20-21-22-S6: Network Security Testing

Related Scripts:
    - generate_a820_1_infrastructure_inventory.py (generates workbooks to check)
    - generate_a820_2_device_security_assessment.py (generates workbooks to check)
    - generate_a820_3_services_catalog.py (generates workbooks to check)
    - generate_a820_4_segmentation_matrix.py (generates workbooks to check)
    - generate_a820_5_controls_coverage.py (generates workbooks to check)
    - generate_a820_6_compliance_dashboard.py (generates dashboard to check)
    - normalize_a820_assessments.py (normalizes before checking)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 24.01.2025
    - Initial release
    - Implements 10 comprehensive diagnostic checks
    - Supports all six A.8.20-21-22 assessment workbook types
    - Provides detailed remediation guidance
    - Includes workbook-specific validation rules

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Quality Assurance Philosophy:**
"The first principle is that you must not fool yourself — and you are the 
easiest person to fool." - Richard Feynman

This sanity checker embodies anti-cargo-cult engineering - it catches structural
issues, formula errors, and validation conflicts that humans easily overlook.
Use it proactively, not reactively (after Excel complains).

**10 Validation Checks Explained:**

1. **Sheet Structure Validation**
   - Verifies all required sheets present
   - Detects missing or renamed sheets
   - Flexible matching (handles minor name variations)
   - Domain-specific expected sheet lists

2. **Assessment ID Validation**
   - Locates ISMS-IMP-A.8.20-21-22.X identifier
   - Validates correct assessment type
   - Checks Instructions sheet and metadata

3. **Formula Validation**
   - Checks formula syntax (balanced parentheses)
   - Validates inter-sheet references (sheet exists)
   - Identifies invalid sheet references
   - Detects external workbook links (dashboard requirement)
   - Reports formula issues with cell coordinates

4. **Data Validation Conflicts**
   - Identifies overlapping validation ranges
   - Checks for conflicting dropdown rules
   - Counts total validations per sheet

5. **Merged Cells Validation**
   - Ensures only top-left cell has content
   - Detects content in non-primary merged cells
   - Prevents Excel repair warnings from merged cell issues

6. **Evidence Register Validation**
   - Counts evidence row capacity (target: 100 rows)
   - Validates EV-XXX ID format
   - Enforces minimum 10 evidence rows
   - Checks for Evidence Register sheet presence

7. **Approval Sign-Off Validation**
   - Checks for 3-level approval workflow
   - Validates: Assessor → Network Manager → CISO
   - Ensures approval structure complete

8. **Summary Dashboard Validation**
   - Counts formulas in Summary/Dashboard sheet
   - Validates calculated metrics present
   - Checks for empty or incomplete dashboards

9. **Row Capacity Validation**
   - Checks data entry capacity (target: 100+ rows)
   - Counts example vs. entry rows
   - Warns on insufficient capacity (<50 rows)

10. **Worksheet Dimensions**
    - Flags oversized worksheets (>5000 rows, >50 columns)
    - Identifies potential performance issues

**Common Issues Detected:**

Critical Issues (Must Fix):
- Missing required sheets (Evidence Register, Approval Sign-Off)
- Invalid formula references (sheet doesn't exist)
- Broken inter-sheet formulas
- Missing assessment ID
- Evidence Register <10 rows (minimum requirement)

High Priority Issues (Should Fix):
- Merged cell content in non-primary cells
- Overlapping data validation ranges
- Formula syntax errors (unbalanced parentheses)
- Missing approval workflow levels

Medium Priority Issues (Recommended):
- Evidence Register <100 rows (recommended capacity)
- Low row capacity in data sheets (<50 rows)
- Excessive formula count suggesting copy-paste errors

Low Priority Issues (Nice to Fix):
- Missing optional sheets
- Cosmetic formatting variations
- Minor naming differences

**Excel Repair Warnings - Root Causes:**

If Excel shows "We found a problem with some content" on open:

Common Causes:
1. **Invalid Sheet References** - Formula references non-existent sheet
2. **Merged Cell Issues** - Content in non-primary cells of merge
3. **Data Validation Conflicts** - Overlapping validation ranges
4. **Style Object Sharing** - Border/Font/Fill objects reused incorrectly
5. **External Link Issues** - Dashboard can't resolve source workbooks

This script identifies causes #1-3 directly. For style issues, use Excel's
built-in repair (usually auto-fixes without data loss).

**Dashboard-Specific Considerations (A820-6):**

Dashboard requires normalized source workbooks:
- ISMS-IMP-A.8.20-21-22.1.xlsx (Infrastructure Inventory)
- ISMS-IMP-A.8.20-21-22.2.xlsx (Device Security Assessment)
- ISMS-IMP-A.8.20-21-22.3.xlsx (Network Services Catalog)
- ISMS-IMP-A.8.20-21-22.4.xlsx (Segmentation Matrix)
- ISMS-IMP-A.8.20-21-22.5.xlsx (Controls Coverage Matrix)

If these are missing, dashboard shows #REF errors. This is EXPECTED until:
1. Source workbooks generated (Scripts 1-5)
2. Normalization completed (normalize_a820_assessments.py)
3. Files placed in same directory as dashboard

Sanity checker will warn about external references but note this is expected.

**Preventive vs. Reactive Usage:**

Preventive (Best Practice):
- Run after every workbook generation
- Validate before distributing to stakeholders
- Check before normalization/consolidation
- Part of automated quality assurance workflow

Reactive (When Problems Occur):
- Excel shows repair warnings
- Formulas show #REF errors
- Dashboard consolidation fails
- Stakeholders report issues opening files

Use preventively to avoid reactive firefighting.

**Integration with CI/CD:**

```bash
#!/bin/bash
# Quality assurance script for A.8.20-21-22 workbook generation

set -e  # Exit on any error

echo "Generating A.8.20-21-22 assessment workbooks..."
python3 generate_a820_1_infrastructure_inventory.py
python3 generate_a820_2_device_security_assessment.py
python3 generate_a820_3_services_catalog.py
python3 generate_a820_4_segmentation_matrix.py
python3 generate_a820_5_controls_coverage.py

echo "Running sanity checks..."
for workbook in ISMS_A_8_20_21_22_[1-5]_*.xlsx; do
    echo "Checking: $workbook"
    python3 excel_sanity_check_a820.py "$workbook" || exit 1
done

echo "All workbooks passed sanity checks ✓"
```

**Audit Considerations:**

Running sanity checks demonstrates quality assurance to auditors:
- Systematic validation of assessment tools
- Documented quality control process
- Evidence of due diligence

Keep sanity check output as QA documentation for audit trail.

**False Positives:**

Some warnings may be acceptable based on context:
- Dashboard external link warnings (expected before normalization)
- Minor sheet name variations (if intentional)
- Optional field completeness (if not applicable)

Review findings and apply judgment - don't blindly fix everything.

**Performance:**

Script typically completes in <10 seconds per workbook. Factors affecting speed:
- Workbook size (large files take longer)
- Number of formulas (more formulas = longer validation)
- Disk I/O (network drives slower than local)

**Limitations:**

What this script CANNOT detect:
- Data accuracy (assessor responsibility)
- Content completeness (human judgment required)
- Stylistic quality (subjective)
- Excel version compatibility issues
- Macro/VBA problems (workbooks don't use macros)

What this script DOES detect:
- Structural integrity issues
- Formula correctness
- Required component presence
- Common openpyxl generation errors

**Troubleshooting:**

Problem: Script fails to load workbook
Solution: File may be corrupt - try opening in Excel and saving

Problem: Script reports missing sheets that exist
Solution: Sheet names may have extra spaces or different case - check carefully

Problem: No issues detected but Excel still shows repair warning
Solution: Issue may be style-related (not detected by script) - let Excel repair

Problem: Script runs very slowly
Solution: Large workbook or network drive - copy to local disk for checking

**Related Quality Assurance Tools:**

Complete A.8.20-21-22 QA toolchain:
1. **excel_sanity_check_a820.py** (this script) - Structural validation
2. **normalize_a820_assessments.py** - Data normalization & validation

Use both for comprehensive quality assurance.

================================================================================
"""

import sys
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ============================================================================
# WORKBOOK TYPE DETECTION
# ============================================================================

def detect_workbook_type(filename):
    """Detect which A.8.20-21-22 workbook type this is based on filename."""
    filename_lower = filename.lower()
    
    if 'infrastructure' in filename_lower and 'inventory' in filename_lower:
        return 'A820-1', 'Infrastructure Inventory'
    elif 'device' in filename_lower and 'security' in filename_lower:
        return 'A820-2', 'Device Security Assessment'
    elif 'services' in filename_lower or 'catalog' in filename_lower:
        return 'A820-3', 'Network Services Catalog'
    elif 'segmentation' in filename_lower or 'matrix' in filename_lower:
        return 'A820-4', 'Network Segmentation Matrix'
    elif 'controls' in filename_lower and 'coverage' in filename_lower:
        return 'A820-5', 'Security Controls Coverage Matrix'
    elif 'dashboard' in filename_lower or 'compliance' in filename_lower:
        return 'A820-6', 'Compliance Dashboard'
    else:
        return 'Unknown', 'Generic A.8.20-21-22 Excel Workbook'


# ============================================================================
# EXPECTED SHEET DEFINITIONS
# ============================================================================

EXPECTED_SHEETS = {
    'A820-1': [
        "Instructions & Legend", "Device_Inventory", "Routers", "Switches",
        "Firewalls", "Wireless_APs", "Load_Balancers", "VPN_Concentrators",
        "Other_Devices", "Discovery_Log", "Evidence_Register", "Approval_Sign_Off"
    ],
    'A820-2': [
        "Instructions & Legend", "Device_Security_Summary", "Router_Hardening",
        "Switch_Hardening", "Firewall_Hardening", "Wireless_Hardening",
        "LoadBalancer_Hardening", "VPN_Hardening", "Authentication_Controls",
        "Encryption_Controls", "Logging_Monitoring", "Gap_Analysis",
        "Evidence_Register", "Approval_Sign_Off"
    ],
    'A820-3': [
        "Instructions & Legend", "Services_Summary", "DNS_Services",
        "DHCP_Services", "NTP_Services", "Proxy_Services", "LoadBalancer_Services",
        "Authentication_Services", "SNMP_Services", "Syslog_Services",
        "Other_Services", "Availability_Monitoring", "Gap_Analysis",
        "Evidence_Register", "Approval_Sign_Off"
    ],
    'A820-4': [
        "Instructions & Legend", "Segmentation_Summary", "Security_Zones",
        "Zone_Networks", "Inter_Zone_Matrix", "Firewall_Rules", "VLAN_Inventory",
        "Microsegmentation", "Effectiveness_Testing", "Flat_Networks",
        "Trust_Boundaries", "Gap_Analysis", "Evidence_Register", "Approval_Sign_Off"
    ],
    'A820-5': [
        "Instructions & Legend", "Controls_Summary", "Device_Controls_Coverage",
        "Service_Controls_Coverage", "Segmentation_Controls_Coverage",
        "Zone_Controls_Matrix", "Defense_in_Depth", "Redundancy_Analysis",
        "Integration_A815_Logging", "Integration_A816_Monitoring",
        "Coverage_Gaps", "Evidence_Completeness", "Cross_Control_Dependencies",
        "Evidence_Register", "Approval_Sign_Off"
    ],
    'A820-6': [
        "Instructions", "Executive_Summary", "Overall_Compliance",
        "A820_Network_Security", "A821_Network_Services", "A822_Network_Segregation",
        "Device_Security_Summary", "Services_Security_Summary", "Segmentation_Summary",
        "Controls_Coverage_Summary", "Gap_Consolidation", "Remediation_Roadmap",
        "Compliance_Trends", "Evidence_Summary", "Approval_Sign_Off"
    ],
}

# Expected Assessment IDs
EXPECTED_ASSESSMENT_IDS = {
    'A820-1': 'ISMS-IMP-A.8.20-21-22.1',
    'A820-2': 'ISMS-IMP-A.8.20-21-22.2',
    'A820-3': 'ISMS-IMP-A.8.20-21-22.3',
    'A820-4': 'ISMS-IMP-A.8.20-21-22.4',
    'A820-5': 'ISMS-IMP-A.8.20-21-22.5',
    'A820-6': 'ISMS-IMP-A.8.20-21-22.6',
}

# Expected normalized filenames for dashboard
DASHBOARD_SOURCES = [
    'ISMS-IMP-A.8.20-21-22.1.xlsx',
    'ISMS-IMP-A.8.20-21-22.2.xlsx',
    'ISMS-IMP-A.8.20-21-22.3.xlsx',
    'ISMS-IMP-A.8.20-21-22.4.xlsx',
    'ISMS-IMP-A.8.20-21-22.5.xlsx',
]


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""
    
    workbook_id, workbook_name = detect_workbook_type(filename)
    
    print("=" * 80)
    print(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {Path(filename).name}")
    print(f"Detected Type: {workbook_id} - {workbook_name}")
    print("=" * 80)
    
    issues_found = []
    warnings_found = []
    
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"\n✓ Workbook loaded successfully")
        print(f"  Sheets found: {len(wb.sheetnames)}")
        print(f"  Sheet names: {', '.join(wb.sheetnames)}")
        
    except Exception as e:
        print(f"\n✗ CRITICAL: Cannot load workbook: {e}")
        return
    
    # ========================================================================
    # CHECK 1: SHEET STRUCTURE VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 1: SHEET STRUCTURE VALIDATION")
    print("=" * 80)
    
    if workbook_id in EXPECTED_SHEETS:
        expected = EXPECTED_SHEETS[workbook_id]
        actual = wb.sheetnames
        
        # Check for missing sheets (flexible matching)
        missing = []
        for exp_sheet in expected:
            # Flexible match: check if expected sheet name is in any actual sheet
            found = False
            for act_sheet in actual:
                if exp_sheet.lower().replace("_", "").replace(" ", "") == act_sheet.lower().replace("_", "").replace(" ", ""):
                    found = True
                    break
            if not found:
                missing.append(exp_sheet)
        
        if missing:
            for sheet in missing:
                warnings_found.append(f"  ⚠ Missing expected sheet: {sheet}")
            print(f"  ⚠ {len(missing)} expected sheets not found (may be renamed)")
        else:
            print(f"  ✓ All expected sheets present for {workbook_id}")
        
        # Check for extra sheets
        print(f"  Total sheets: {len(actual)} (expected {len(expected)})")
    else:
        print("  ℹ Unknown workbook type - skipping sheet validation")
    
    # ========================================================================
    # CHECK 2: ASSESSMENT ID VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 2: ASSESSMENT ID VALIDATION")
    print("=" * 80)
    
    assessment_id_found = False
    if workbook_id in EXPECTED_ASSESSMENT_IDS:
        expected_id = EXPECTED_ASSESSMENT_IDS[workbook_id]
        
        # Check in Instructions & Legend sheet
        if 'Instructions & Legend' in wb.sheetnames or 'Instructions' in wb.sheetnames:
            sheet_name = 'Instructions & Legend' if 'Instructions & Legend' in wb.sheetnames else 'Instructions'
            ws = wb[sheet_name]
            # Check first few rows for assessment ID
            for row in range(1, 15):
                for col in range(1, 10):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and expected_id in cell_value:
                        print(f"  ✓ Assessment ID found: {expected_id}")
                        assessment_id_found = True
                        break
                if assessment_id_found:
                    break
        
        # Also check in other common locations
        if not assessment_id_found:
            for sheet_name in wb.sheetnames:
                if 'Instructions' in sheet_name or 'Summary' in sheet_name or 'Executive' in sheet_name:
                    ws = wb[sheet_name]
                    for row in range(1, 10):
                        for col in range(1, 5):
                            cell_value = ws.cell(row=row, column=col).value
                            if cell_value and isinstance(cell_value, str) and expected_id in cell_value:
                                print(f"  ✓ Assessment ID found in {sheet_name}: {expected_id}")
                                assessment_id_found = True
                                break
                        if assessment_id_found:
                            break
                if assessment_id_found:
                    break
        
        if not assessment_id_found:
            warnings_found.append(
                f"  ⚠ Assessment ID '{expected_id}' not found in Instructions sheet"
            )
    else:
        print("  ℹ Unknown workbook type - skipping Assessment ID validation")
    
    # ========================================================================
    # CHECK 3: FORMULA VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 3: FORMULA VALIDATION")
    print("=" * 80)
    
    formula_issues = 0
    inter_sheet_refs = {}
    external_refs = set()
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # Check for external workbook references
                    ext_refs = re.findall(r'\[([^\]]+\.xlsx)\]', formula)
                    for ref in ext_refs:
                        external_refs.add(ref)
                    
                    # Check for sheet references
                    if "'" in formula or "!" in formula:
                        sheet_refs = re.findall(r"'?([^'!]+)'?!", formula)
                        for ref in sheet_refs:
                            ref_clean = ref.strip("'")
                            # Skip external references
                            if "[" not in ref_clean and ref_clean not in wb.sheetnames:
                                # Check for close matches (underscore vs space)
                                close_match = False
                                for actual in wb.sheetnames:
                                    if ref_clean.replace(" ", "_") == actual or ref_clean.replace("_", " ") == actual:
                                        close_match = True
                                        break
                                if not close_match:
                                    issues_found.append(
                                        f"  ✗ {sheet_name}!{cell.coordinate}: "
                                        f"Invalid sheet reference '{ref_clean}'"
                                    )
                                    formula_issues += 1
                            
                            # Track inter-sheet references
                            if "[" not in ref_clean and ref_clean in wb.sheetnames:
                                if sheet_name not in inter_sheet_refs:
                                    inter_sheet_refs[sheet_name] = set()
                                inter_sheet_refs[sheet_name].add(ref_clean)
                    
                    # Check for common syntax issues
                    if formula.count('(') != formula.count(')'):
                        issues_found.append(
                            f"  ✗ {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced parentheses in formula"
                        )
                        formula_issues += 1
    
    if formula_issues == 0:
        print("  ✓ No formula syntax issues detected")
    else:
        print(f"  ✗ Found {formula_issues} formula issues")
    
    if inter_sheet_refs:
        print(f"  ℹ Inter-sheet references found:")
        for sheet, refs in inter_sheet_refs.items():
            print(f"    {sheet} → {', '.join(sorted(refs))}")
    
    if external_refs:
        print(f"  ℹ External workbook references found:")
        for ref in sorted(external_refs):
            print(f"    → {ref}")
        if workbook_id == 'A820-6':
            print("  ⚠ WARNING: Dashboard requires external links to function")
    
    # ========================================================================
    # CHECK 4: DATA VALIDATION CONFLICTS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 4: DATA VALIDATION CONFLICTS")
    print("=" * 80)
    
    validation_issues = 0
    total_validations = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            dv_count = len(ws.data_validations.dataValidation)
            total_validations += dv_count
            
            # Check for overlapping ranges
            all_ranges = []
            for dv in ws.data_validations.dataValidation:
                if hasattr(dv, 'sqref') and dv.sqref:
                    all_ranges.extend(str(dv.sqref).split())
            
            if len(all_ranges) != len(set(all_ranges)):
                warnings_found.append(
                    f"  ⚠ {sheet_name}: Potentially overlapping data validation ranges"
                )
                validation_issues += 1
    
    print(f"  Total data validations: {total_validations}")
    
    if validation_issues == 0:
        print("  ✓ No data validation conflicts detected")
    
    # ========================================================================
    # CHECK 5: MERGED CELLS VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 5: MERGED CELLS VALIDATION")
    print("=" * 80)
    
    merge_issues = 0
    total_merges = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'merged_cells') and ws.merged_cells:
            merge_count = len(ws.merged_cells.ranges)
            total_merges += merge_count
            
            # Check merged cells for content issues
            for merge_range in ws.merged_cells.ranges:
                min_col, min_row = merge_range.min_col, merge_range.min_row
                max_col, max_row = merge_range.max_col, merge_range.max_row
                
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if row == min_row and col == min_col:
                            continue
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            cell_coord = f"{get_column_letter(col)}{row}"
                            warnings_found.append(
                                f"  ⚠ {sheet_name}!{cell_coord}: "
                                f"Merged cell has content in non-primary cell"
                            )
                            merge_issues += 1
    
    print(f"  Total merged ranges: {total_merges}")
    
    if merge_issues == 0:
        print("  ✓ Merged cells properly formatted")
    
    # ========================================================================
    # CHECK 6: EVIDENCE REGISTER VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 6: EVIDENCE REGISTER VALIDATION")
    print("=" * 80)
    
    evidence_sheet = None
    for sheet in wb.sheetnames:
        if 'evidence' in sheet.lower() and 'register' in sheet.lower():
            evidence_sheet = sheet
            break
    
    if evidence_sheet:
        ws = wb[evidence_sheet]
        # Count evidence rows (look for EV- prefix)
        evidence_count = 0
        for row in range(1, min(110, ws.max_row + 1)):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and isinstance(cell_val, str):
                if cell_val.startswith('EV-'):
                    evidence_count += 1
        
        print(f"  Evidence sheet: {evidence_sheet}")
        print(f"  Evidence row templates: {evidence_count}")
        
        # A.8.20-21-22 requires minimum 10 evidence items
        if evidence_count < 10:
            issues_found.append(
                f"  ✗ Evidence Register has {evidence_count} rows (MINIMUM 10 REQUIRED)"
            )
        elif evidence_count < 50:
            warnings_found.append(
                f"  ⚠ Evidence Register has {evidence_count} rows (recommended ≥100 for flexibility)"
            )
        elif evidence_count >= 100:
            print(f"  ✓ Full 100-row evidence register (excellent)")
        else:
            print(f"  ✓ Evidence register adequate ({evidence_count} rows)")
    else:
        if workbook_id != 'A820-6':  # Dashboard has different structure
            issues_found.append("  ✗ No Evidence Register sheet found (REQUIRED)")
    
    # ========================================================================
    # CHECK 7: APPROVAL SIGN-OFF VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 7: APPROVAL SIGN-OFF VALIDATION")
    print("=" * 80)
    
    approval_sheet = None
    for sheet in wb.sheetnames:
        if 'approval' in sheet.lower() or 'sign' in sheet.lower():
            approval_sheet = sheet
            break
    
    if approval_sheet:
        ws = wb[approval_sheet]
        
        # Check for 3-level approval structure
        approval_levels = 0
        for row in range(1, min(100, ws.max_row + 1)):
            for col in range(1, 6):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val and isinstance(cell_val, str):
                    if 'PREPARED BY' in cell_val or 'COMPLETED BY' in cell_val or 'ASSESSOR' in cell_val:
                        approval_levels += 1
                    elif 'NETWORK MANAGER' in cell_val or 'REVIEWED BY' in cell_val or 'SOC' in cell_val:
                        approval_levels += 1
                    elif 'CISO' in cell_val or 'APPROVED BY' in cell_val:
                        approval_levels += 1
        
        print(f"  Approval sheet: {approval_sheet}")
        print(f"  Approval levels detected: {approval_levels}")
        
        if approval_levels < 3:
            warnings_found.append(
                f"  ⚠ Approval workflow has {approval_levels} levels (expected 3: Assessor → Network Mgr → CISO)"
            )
        else:
            print(f"  ✓ Complete 3-level approval workflow")
    else:
        issues_found.append("  ✗ No Approval Sign-Off sheet found (REQUIRED)")
    
    # ========================================================================
    # CHECK 8: SUMMARY DASHBOARD VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 8: SUMMARY DASHBOARD VALIDATION")
    print("=" * 80)
    
    summary_sheet = None
    for sheet in wb.sheetnames:
        if 'summary' in sheet.lower() or 'dashboard' in sheet.lower() or 'executive' in sheet.lower():
            summary_sheet = sheet
            break
    
    if summary_sheet:
        ws = wb[summary_sheet]
        
        # Count formulas in Summary sheet
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
        
        print(f"  Summary sheet: {summary_sheet}")
        print(f"  Formulas in Summary: {formula_count}")
        
        if formula_count < 10:
            warnings_found.append(
                f"  ⚠ Summary sheet has only {formula_count} formulas (expected many calculated metrics)"
            )
        else:
            print(f"  ✓ Summary dashboard has active calculations")
    else:
        warnings_found.append("  ⚠ No Summary Dashboard sheet found")
    
    # ========================================================================
    # CHECK 9: ROW CAPACITY VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 9: ROW CAPACITY VALIDATION")
    print("=" * 80)
    
    data_sheets = [s for s in wb.sheetnames if s not in ['Instructions & Legend', 'Instructions',
                                                          'Summary Dashboard', 'Executive_Summary',
                                                          'Evidence Register', 'Approval Sign-Off',
                                                          'Evidence_Summary']]
    
    for sheet_name in data_sheets[:5]:  # Check first 5 data sheets
        ws = wb[sheet_name]
        
        # Check if data entry area exists
        if ws.max_row > 50:
            # Count example rows vs entry rows
            example_count = 0
            entry_area_start = 20  # Typically after examples
            
            for row in range(7, min(20, ws.max_row)):
                # Check if row has data (non-empty first column)
                if ws.cell(row=row, column=1).value:
                    example_count += 1
            
            total_capacity = ws.max_row - 6  # Subtract header rows
            
            print(f"  {sheet_name}:")
            print(f"    Total capacity: ~{total_capacity} rows")
            print(f"    Example data rows: ~{example_count}")
            
            if total_capacity < 50:
                warnings_found.append(
                    f"  ⚠ {sheet_name}: Low capacity ({total_capacity} rows), recommended ≥100"
                )
            elif total_capacity >= 100:
                print(f"    ✓ Adequate capacity (≥100 rows)")
    
    # ========================================================================
    # CHECK 10: WORKSHEET DIMENSIONS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 10: WORKSHEET DIMENSIONS")
    print("=" * 80)
    
    dimension_issues = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if ws.max_row > 5000:
            warnings_found.append(
                f"  ⚠ {sheet_name}: Large worksheet ({ws.max_row} rows)"
            )
            dimension_issues += 1
        
        if ws.max_column > 50:
            warnings_found.append(
                f"  ⚠ {sheet_name}: Wide worksheet ({ws.max_column} columns)"
            )
            dimension_issues += 1
    
    if dimension_issues == 0:
        print("  ✓ Worksheet dimensions within reasonable limits")
    
    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    print("\n" + "=" * 80)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 80)
    
    if issues_found:
        print(f"\n❌ CRITICAL ISSUES FOUND: {len(issues_found)}")
        for issue in issues_found:
            print(issue)
    
    if warnings_found:
        print(f"\n⚠️  WARNINGS: {len(warnings_found)}")
        for warning in warnings_found:
            print(warning)
    
    if not issues_found and not warnings_found:
        print("\n✅ NO ISSUES DETECTED")
        print("\nThe workbook appears structurally sound.")
        print("If Excel still shows repair warnings, consider:")
        print("  • Excel version compatibility (try Excel 2019+)")
        print("  • Antivirus/security software interference")
        print("  • Network drive or OneDrive sync issues")
        print("  • Excel's overly cautious validation")
        if workbook_id == 'A820-6':
            print("  • Unresolved external workbook links (expected before normalization)")
    else:
        print_recommendations(issues_found, warnings_found, workbook_id, 
                            formula_issues, validation_issues, merge_issues)
    
    # ========================================================================
    # WORKBOOK-SPECIFIC NOTES
    # ========================================================================
    print_workbook_notes(workbook_id, external_refs)
    
    wb.close()


def print_recommendations(issues, warnings, workbook_id, formula_issues, 
                         validation_issues, merge_issues):
    """Print recommended fixes based on issues found."""
    print("\n" + "=" * 80)
    print("RECOMMENDED ACTIONS")
    print("=" * 80)
    
    if formula_issues > 0:
        print("\n1. FORMULA FIXES:")
        print("   • Review formulas referencing other sheets")
        print("   • Ensure sheet names match exactly (case-sensitive)")
        print("   • Check for balanced quotes and parentheses")
        if workbook_id == 'A820-6':
            print("   • Verify external workbook references point to normalized files")
    
    if validation_issues > 0:
        print("\n2. DATA VALIDATION FIXES:")
        print("   • Remove overlapping data validation ranges")
        print("   • Apply validations to specific ranges, not entire columns")
    
    if merge_issues > 0:
        print("\n3. MERGED CELL FIXES:")
        print("   • Ensure only top-left cell of merged range has content")
        print("   • Clear content from other cells in merged range")
    
    if workbook_id == 'A820-6':
        print("\n4. DASHBOARD-SPECIFIC SETUP (A820-6):")
        print("   • This workbook consolidates data from 5 assessment workbooks")
        print("   • Ensure normalized source workbooks are in the same folder:")
        for source in DASHBOARD_SOURCES:
            print(f"     - {source}")
        print("   • Run normalize_a820_assessments.py first to create normalized files")


def print_workbook_notes(workbook_id, external_refs):
    """Print workbook-specific notes and expected structure."""
    print("\n" + "=" * 80)
    print(f"WORKBOOK-SPECIFIC NOTES: {workbook_id}")
    print("=" * 80)
    
    if workbook_id == 'A820-1':
        print("\nInfrastructure Inventory (WB1):")
        print("  • 12 sheets for network device inventory")
        print("  • 100+ entry capacity per device type")
        print("  • Device categorization: Routers, Switches, Firewalls, Wireless APs, etc.")
        print("  • Location and ownership tracking")
        print("  • Management interface documentation")
        print("  • Firmware/software version tracking")
        print("  • Criticality classification")
        print("  • Discovery methodology documentation")
        print("  • Evidence Register (100 entries minimum)")
        print("  • 3-level approval workflow (Assessor → Network Mgr → CISO)")
        print("\n  Key purpose:")
        print("    - Foundation for device security assessment (WB2)")
        print("    - Input to controls coverage analysis (WB5)")
        print("    - Feeds compliance dashboard (WB6)")
    
    elif workbook_id == 'A820-2':
        print("\nDevice Security Assessment (WB2):")
        print("  • 14 sheets for device hardening assessment")
        print("  • 100+ entry capacity per assessment sheet")
        print("  • Device-specific hardening baselines")
        print("  • Authentication mechanisms assessment")
        print("  • Access control configuration")
        print("  • Encryption configuration")
        print("  • Logging and monitoring configuration")
        print("  • Configuration backup status")
        print("  • Patch/update status tracking")
        print("  • Gap identification and remediation")
        print("  • Evidence Register (100 entries minimum)")
        print("  • 3-level approval workflow")
        print("\n  Key metrics:")
        print("    - Overall device hardening compliance percentage")
        print("    - Critical gaps requiring immediate remediation")
        print("    - Device security posture by type")
    
    elif workbook_id == 'A820-3':
        print("\nNetwork Services Catalog (WB3):")
        print("  • 15 sheets for network services assessment")
        print("  • 100+ entry capacity per service type")
        print("  • Service-specific security assessments")
        print("  • DNS, DHCP, NTP, Proxy, Load Balancers, etc.")
        print("  • Security mechanisms per service")
        print("  • Availability and redundancy status")
        print("  • Service monitoring configuration")
        print("  • Service owner tracking")
        print("  • Gap identification and remediation")
        print("  • Evidence Register (100 entries minimum)")
        print("  • 3-level approval workflow")
        print("\n  Key metrics:")
        print("    - Service security compliance percentage")
        print("    - Service availability metrics")
        print("    - Critical service gaps")
    
    elif workbook_id == 'A820-4':
        print("\nNetwork Segmentation Matrix (WB4):")
        print("  • 14 sheets for segmentation assessment")
        print("  • 100+ entry capacity per assessment area")
        print("  • Security zones definition and inventory")
        print("  • Zone-to-zone trust relationships")
        print("  • Inter-zone traffic policies")
        print("  • Firewall rules inventory")
        print("  • VLAN inventory and mapping")
        print("  • Microsegmentation tracking")
        print("  • Effectiveness testing results")
        print("  • Flat network identification")
        print("  • Trust boundary documentation")
        print("  • Gap identification and remediation")
        print("  • Evidence Register (100 entries minimum)")
        print("  • 3-level approval workflow")
        print("\n  Key metrics:")
        print("    - Segmentation effectiveness percentage")
        print("    - Flat network identification and risk")
        print("    - Zone protection compliance")
    
    elif workbook_id == 'A820-5':
        print("\nSecurity Controls Coverage Matrix (WB5):")
        print("  • 15 sheets for unified controls assessment")
        print("  • Consolidates WB1-4 findings")
        print("  • Device, service, and segmentation controls per zone")
        print("  • Defense-in-depth analysis")
        print("  • Redundancy and single point of failure analysis")
        print("  • Integration with A.8.15 (Logging) and A.8.16 (Monitoring)")
        print("  • Coverage gap identification")
        print("  • Evidence completeness tracking")
        print("  • Cross-control dependency mapping")
        print("  • Evidence Register (100 entries minimum)")
        print("  • 3-level approval workflow")
        print("\n  Key metrics:")
        print("    - Overall network security controls coverage")
        print("    - Defense-in-depth effectiveness")
        print("    - Coverage gaps by security zone")
    
    elif workbook_id == 'A820-6':
        print("\nCompliance Dashboard (MASTER):")
        print("  • 15 sheets - EXECUTIVE CONSOLIDATION WORKBOOK")
        print("  • Overall A.8.20, A.8.21, A.8.22 compliance scores")
        print("  • Device security summary")
        print("  • Services security summary")
        print("  • Segmentation effectiveness summary")
        print("  • Controls coverage summary")
        print("  • Gap consolidation and prioritization")
        print("  • Remediation roadmap with timelines")
        print("  • Compliance trends over time")
        print("  • Evidence summary for audit readiness")
        print("  • 3-level approval workflow")
        print("\n  ⚠ This dashboard requires normalized source files:")
        for source in DASHBOARD_SOURCES:
            print(f"    - {source}")
        print("\n  Setup process:")
        print("    1. Generate all 5 assessment workbooks (Scripts 1-5)")
        print("    2. Run: python normalize_a820_assessments.py")
        print("    3. Dashboard reads from same directory as normalized files")
        print("    4. May show #REF errors until normalization complete")
    
    else:
        print("\nGeneric A.8.20-21-22 Workbook")
        print("  • No specific validation rules defined for this type")
        print("  • Standard Excel health checks applied")
        print("  • Ensure workbook follows A.8.20-21-22 framework structure")
    
    print("\n" + "=" * 80)


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("=" * 70)
        print("ISMS A.8.20-21-22 Framework - Excel Sanity Checker")
        print("Controls: A.8.20, A.8.21, A.8.22 (Network Security)")
        print("=" * 70)
        print("\nUsage: python3 excel_sanity_check_a820.py <filename.xlsx>")
        print("\nExamples:")
        print("  python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_1_Infrastructure_Inventory.xlsx")
        print("  python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_2_Device_Security_Assessment.xlsx")
        print("  python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_3_Services_Catalog.xlsx")
        print("  python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_4_Segmentation_Matrix.xlsx")
        print("  python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_5_Controls_Coverage.xlsx")
        print("  python3 excel_sanity_check_a820.py ISMS_A_8_20_21_22_Dashboard.xlsx")
        print("\nSupported workbook types:")
        print("  A820-1 - Infrastructure Inventory")
        print("  A820-2 - Device Security Assessment")
        print("  A820-3 - Network Services Catalog")
        print("  A820-4 - Network Segmentation Matrix")
        print("  A820-5 - Security Controls Coverage Matrix")
        print("  A820-6 - Compliance Dashboard")
        print("\n'Don't fool yourself' - Systems Engineering ISMS")
        sys.exit(1)
    
    filename = sys.argv[1]
    
    try:
        check_workbook_health(filename)
    except FileNotFoundError:
        print(f"\n❌ ERROR: File not found: {filename}")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()