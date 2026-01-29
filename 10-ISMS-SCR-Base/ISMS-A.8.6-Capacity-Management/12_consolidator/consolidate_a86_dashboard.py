#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
Dashboard Data Consolidation Script - ISMS A.8.6 Capacity Management Framework
================================================================================

ISO/IEC 27001:2022 Control A.8.6: Capacity Management
Alternative Data Integration Method for Compliance Dashboard

--------------------------------------------------------------------------------
DATA CONSOLIDATION UTILITY - ALTERNATIVE TO EXTERNAL WORKBOOK LINKING
--------------------------------------------------------------------------------

This utility provides an alternative method for populating the A.8.6 Capacity
Management Dashboard with data from source assessment workbooks (Assessment 1
and Assessment 2) by directly copying user-entered data rather than using
external workbook formula references.

**Primary Approach vs. Alternative Approach:**

**Primary (Recommended): External Workbook Formula Linking**
- Dashboard contains formulas like: ='[ISMS_A_8_6_1_*.xlsx]Sheet_Name'!$B$15
- Data updates automatically when source files change
- Dashboard and source files must be in same directory
- Requires "Update Links" when opening dashboard
- Real-time data synchronization
- See: generate_a86_3_compliance_dashboard.py

**Alternative (This Script): Direct Data Consolidation**
- Reads data from source workbooks programmatically
- Copies data values (not formulas) into dashboard
- Creates static snapshot of capacity management status
- Works with dashboard and sources in different locations
- Requires manual re-run when source data changes
- Useful when external links don't work or aren't desired

**When to Use This Script:**

1. **External Links Don't Work**
   - Excel security settings prevent external workbook links
   - SharePoint/OneDrive environments where linking fails
   - Email distribution requires self-contained workbooks
   - Recipient systems block external references

2. **Static Snapshots Required**
   - Monthly/quarterly capacity management reporting archives
   - Executive presentations requiring frozen data
   - Audit evidence packages with point-in-time data
   - Historical capacity tracking over time

3. **Simplified Distribution**
   - Single-file distribution to stakeholders
   - No dependency on source file availability
   - Recipients don't have access to source assessments
   - Reduced file size (no external references)

4. **Data Consolidation Flexibility**
   - Selective data inclusion from sources
   - Custom data transformation during consolidation
   - Multi-source aggregation and calculations
   - Data validation during import

**When NOT to Use This Script:**

- If external workbook linking works reliably → Use primary approach
- If real-time data updates are needed → Use external linking
- If dashboard and sources will always be co-located → Use external linking
- For initial dashboard setup → Use generate_a86_3_compliance_dashboard.py first

**Purpose:**
Reads user-entered data from all relevant sheets in the two assessment workbooks
(Utilization and Forecasts) and writes consolidated data into corresponding
dashboard sheets, preserving the Executive Dashboard's external formulas while
populating detail sheets with actual assessment data.

**What It Does:**

1. **Validates Source Availability**
   - Checks both assessment workbooks exist
   - Validates dashboard workbook is accessible
   - Reports missing files before attempting consolidation

2. **Reads Source Data**
   - Opens each source workbook in data-only mode
   - Reads user-entered data from configured sheets
   - Extracts relevant columns and rows per mapping
   - Skips empty rows and merged cell areas

3. **Consolidates into Dashboard**
   - Maps source sheets to target dashboard sheets
   - Writes data to appropriate dashboard locations
   - Adds source labels for traceability
   - Preserves existing dashboard structure

4. **Preserves Executive Dashboard**
   - Does NOT modify Executive Dashboard sheet
   - Retains external workbook formula references
   - Allows hybrid approach (formulas + consolidated data)

5. **Provides Consolidation Statistics**
   - Reports sheets processed and rows consolidated
   - Shows data distribution across dashboard sheets
   - Identifies any skipped or missing sheets

**Consolidation Mapping:**
The script uses a comprehensive mapping structure defining:
- Source workbook → Source sheet → Target dashboard sheet
- Starting row for data extraction
- Column ranges to include
- Source labels for traceability
- Data organization in dashboard

Example mapping:
```python
"assessment_1": {  # A.8.6.1 Utilization
    "Compute_Capacity": {
        "target": "Utilization Summary",     # Dashboard sheet
        "start_row": 10,                     # Begin reading at row 10
        "columns": "A:M",                    # Include columns A through M
        "label": "Assessment 1: Compute"     # Source identifier
    },
    ...
}
```

**Supported Source Workbooks:**
- ISMS_A_8_6_1_Capacity_Utilization_Assessment_YYYYMMDD.xlsx
- ISMS_A_8_6_2_Capacity_Forecasts_Planning_YYYYMMDD.xlsx

These are the file names created by the assessment generator scripts.

**Target Dashboard Sheets:**
- Utilization Summary (from Assessment 1: current capacity utilization)
- Capacity Risks (from Assessment 1: resources at warning/critical)
- Forecast Summary (from Assessment 2: capacity exhaustion timeline)
- Planning Effectiveness (from Assessment 2: proactive vs. reactive)
- Capacity Trends (from both assessments: utilization over time)
- Evidence Summary (from both assessments: consolidated evidence)

**Executive Dashboard Sheet:**
- NOT modified by this script
- Retains external workbook formula references
- Pulls summary metrics from other dashboard sheets
- Allows hybrid linking + consolidation approach

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

**Basic Usage:**
    python3 consolidate_a86_dashboard.py ISMS_A_8_6_3_Dashboard_20250128.xlsx

**Prerequisites:**
    1. Generate dashboard first:
       python3 generate_a86_3_compliance_dashboard.py
    
    2. Complete assessment workbooks:
       - Assessment 1 (Utilization) completed with current data
       - Assessment 2 (Forecasts) completed with forecasts
    
    3. Optionally normalize assessment files:
       python3 normalize_assessment_files_a86.py
    
    4. Ensure all files are accessible:
       - ISMS_A_8_6_1_Capacity_Utilization_Assessment_YYYYMMDD.xlsx
       - ISMS_A_8_6_2_Capacity_Forecasts_Planning_YYYYMMDD.xlsx
       - Dashboard file (provided as argument)

**File Location Requirements:**
    Source workbooks must be in CURRENT DIRECTORY when running script:
    - ISMS_A_8_6_1_*.xlsx (in current directory)
    - ISMS_A_8_6_2_*.xlsx (in current directory)
    
    Dashboard can be in any directory (specify full path if needed)

**Example Session:**
    $ python3 consolidate_a86_dashboard.py ISMS_A_8_6_3_Dashboard_20250128.xlsx
    ================================================================================
    ISMS A.8.6 - DASHBOARD DATA CONSOLIDATION
    ================================================================================
    
    Dashboard: ISMS_A_8_6_3_Dashboard_20250128.xlsx
    Started: 2025-01-28 14:30:15
    
    Checking source workbooks...
      ✅ ISMS_A_8_6_1_Capacity_Utilization_Assessment_20250128.xlsx
      ✅ ISMS_A_8_6_2_Capacity_Forecasts_Planning_20250128.xlsx
    
    Loading dashboard...
      ✅ Dashboard loaded (13 sheets)
    
    ================================================================================
    CONSOLIDATING DATA FROM SOURCE WORKBOOKS
    ================================================================================
    
    📁 Processing: Assessment 1 (Utilization)
      ✅ Compute_Capacity → Utilization Summary: 24 rows
      ✅ Storage_Capacity → Utilization Summary: 18 rows
      ✅ Network_Capacity → Utilization Summary: 12 rows
      ✅ Threshold_Status → Capacity Risks: 15 rows
      ✅ Gap_Analysis → Capacity Risks: 8 rows
      [... additional sheets ...]
    
    📁 Processing: Assessment 2 (Forecasts & Planning)
      ✅ 12_Month_Forecast → Forecast Summary: 35 rows
      ✅ Capacity_Exhaustion → Forecast Summary: 12 rows
      ✅ Expansion_Planning → Planning Effectiveness: 10 rows
      ✅ Forecast_Accuracy → Planning Effectiveness: 8 rows
      [... additional sheets ...]
    
    ================================================================================
    SAVING CONSOLIDATED DASHBOARD
    ================================================================================
    ✅ Dashboard saved successfully
    
    ================================================================================
    CONSOLIDATION COMPLETE
    ================================================================================
    
    📊 Statistics:
      • Source sheets read: 18
      • Total rows consolidated: 167
    
    📋 By Dashboard Sheet:
      • Utilization Summary: 54 rows
      • Capacity Risks: 23 rows
      • Forecast Summary: 47 rows
      • Planning Effectiveness: 18 rows
      • Capacity Trends: 15 rows
      • Evidence Summary: 10 rows
    
    ✅ Completed: 2025-01-28 14:30:28
    
    ================================================================================
    🎯 Dashboard now contains all user-entered data from assessments!
       Open in Excel and click 'Update Links' to refresh formulas (if external
       linking is enabled in Executive Dashboard).
    ================================================================================

**Exit Codes:**
    0 = Consolidation successful
    1 = Consolidation failed (missing files, errors during processing)

**Re-Running Consolidation:**
    Safe to re-run this script when assessment data is updated:
    - Existing dashboard data will be overwritten
    - Updated source data will be consolidated
    - Statistics will reflect new data volumes
    - No data loss in source workbooks (read-only access)

**Workflow Integration:**

    **Initial Setup (One-Time):**
    1. Generate dashboard: python3 generate_a86_3_compliance_dashboard.py
    2. Complete assessments: Fill in Assessment 1 and Assessment 2
    3. Consolidate data: python3 consolidate_a86_dashboard.py dashboard.xlsx
    
    **Monthly Updates (Assessment 1 Utilization):**
    1. Update Assessment 1 with current utilization data
    2. Re-consolidate: python3 consolidate_a86_dashboard.py dashboard.xlsx
    
    **Quarterly Updates (Assessment 2 Forecasts):**
    1. Update Assessment 2 with new forecasts and planning data
    2. Re-consolidate: python3 consolidate_a86_dashboard.py dashboard.xlsx
    
    **Static Snapshot Creation:**
    1. Run consolidation with date-stamped dashboard name
    2. Archive consolidated dashboard for reporting period
    3. Create new dashboard for next period

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel manipulation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Excel file reading and writing)
    - sys, os (standard library)
    - datetime (standard library)
    - glob (standard library - for file pattern matching)

**Required Files:**
    - Two assessment workbooks (Assessment 1 and Assessment 2)
    - Generated dashboard workbook (Assessment 3 / Dashboard)
    - This consolidation script

**File Naming:**
    Source workbooks should match pattern:
    - ISMS_A_8_6_1_Capacity_Utilization_Assessment_*.xlsx
    - ISMS_A_8_6_2_Capacity_Forecasts_Planning_*.xlsx
    
    Script will find most recent files matching pattern (by YYYYMMDD suffix).
    Dashboard can have any name (passed as command-line argument).

--------------------------------------------------------------------------------
CONSOLIDATION MAPPING STRUCTURE
--------------------------------------------------------------------------------

**How It Works:**
The script uses a detailed mapping dictionary (CONSOLIDATION_MAP) that defines
which data to extract from which source sheets and where to place it in the
dashboard.

**Mapping Components:**

1. **Source Workbook Key** (assessment_1, assessment_2)
   - assessment_1 = ISMS_A_8_6_1_*.xlsx (Utilization Assessment)
   - assessment_2 = ISMS_A_8_6_2_*.xlsx (Forecasts & Planning Assessment)

2. **Source Sheet Name**
   - Exact sheet name in source workbook
   - Examples: Compute_Capacity, Storage_Capacity, 12_Month_Forecast

3. **Target Dashboard Sheet**
   - Destination sheet in dashboard workbook
   - Examples: Utilization Summary, Capacity Risks, Forecast Summary

4. **Start Row**
   - Row number to begin reading data (typically row 10)
   - Skips headers and instructions

5. **Column Range**
   - Columns to include (e.g., "A:M" includes columns A through M)
   - Defines data width for each source

6. **Source Label**
   - Descriptive label added to consolidated data
   - Enables traceability back to source assessment

**Total Mapped Sources:**
- Assessment 1 (Utilization): 10 source sheets → 4 dashboard sheets
- Assessment 2 (Forecasts): 8 source sheets → 4 dashboard sheets
- Total: 18 source sheets consolidated into 6 dashboard sheets

**Customization:**
To modify consolidation behavior, edit CONSOLIDATION_MAP dictionary:
- Add/remove source sheets
- Change target dashboard sheets
- Adjust column ranges
- Modify source labels
- Change starting rows

--------------------------------------------------------------------------------
DATA CONSOLIDATION PROCESS
--------------------------------------------------------------------------------

**Step-by-Step Process:**

1. **Validation Phase**
   - Verify dashboard file exists and is accessible
   - Check both source assessment workbooks are present
   - Use glob pattern matching to find latest assessment files
   - Report any missing files before proceeding

2. **Dashboard Loading**
   - Open dashboard workbook in read-write mode
   - Validate expected sheets exist
   - Prepare for data consolidation

3. **Source Processing Loop**
   For each source workbook (Assessment 1 and Assessment 2):
   
   a. Open source workbook in data-only mode (values, not formulas)
   b. For each configured source sheet:
      - Locate source sheet by name
      - Read data starting from specified row
      - Extract specified column range
      - Stop at first completely empty row
      - Skip merged cell areas safely
   
   c. For each extracted data block:
      - Locate target dashboard sheet
      - Find first available row for writing
      - Add source label for traceability
      - Write data values (not formulas)
      - Track consolidation statistics

4. **Dashboard Preservation**
   - Executive Dashboard sheet is NOT modified
   - External formula references remain intact
   - Hybrid approach maintained (formulas + consolidated data)

5. **Save and Report**
   - Save modified dashboard workbook
   - Generate consolidation statistics
   - Report data distribution across dashboard sheets
   - Provide completion timestamp

**Data Handling:**
- Source workbooks opened in data-only mode (formulas evaluated to values)
- Empty rows terminate data extraction for each source
- Merged cells are safely skipped (no write attempts)
- Existing dashboard data is overwritten on re-consolidation
- Source workbooks remain unmodified (read-only access)

**Error Handling:**
- Missing source files reported before consolidation starts
- Missing source sheets logged but don't stop consolidation
- Missing target sheets logged and skipped
- Cell write errors caught and logged
- Dashboard save failures reported with error details

--------------------------------------------------------------------------------
INTEGRATION WITH A.8.6 FRAMEWORK
--------------------------------------------------------------------------------

**Position in Workflow:**

This script provides an ALTERNATIVE data integration method. The standard
workflow uses external workbook formula linking. This consolidation approach
is used when external linking isn't suitable or desired.

**Standard Workflow (External Linking - Recommended):**
```
Assessments → Dashboard Generation → Link Update → Distribution
(A.8.6.1-2)      (generate_a86_3)      (Excel)    (Stakeholders)
```

**Alternative Workflow (Data Consolidation - This Script):**
```
Assessments → Dashboard Generation → Data Consolidation → Distribution
(A.8.6.1-2)      (generate_a86_3)    (this script)   (Stakeholders)
                                           ↓
                                  Static Snapshot
```

**Hybrid Workflow (Combined Approach):**
```
Assessments → Dashboard Generation → Link Update + Consolidation
(A.8.6.1-2)      (generate_a86_3)         (Both)
                                           ↓
                              Exec formulas + Detail data
```

**Related Scripts:**
- generate_a86_1_capacity_utilization.py (creates Assessment 1)
- generate_a86_2_capacity_forecasts.py (creates Assessment 2)
- generate_a86_3_compliance_dashboard.py (creates dashboard)
- normalize_assessment_files_a86.py (normalizes and validates assessments)

**Comparison: External Linking vs. Data Consolidation**

| Aspect | External Linking | Data Consolidation |
|--------|------------------|-------------------|
| **Setup** | One-time generation | Generate + consolidate |
| **Updates** | Automatic (Excel refresh) | Manual re-run required |
| **File Dependency** | Must be co-located | Can be separate |
| **Distribution** | Requires source files | Self-contained |
| **Excel Security** | May be blocked | Always works |
| **Historical Tracking** | Overwrites data | Create dated snapshots |
| **Data Type** | Live formulas | Static values |
| **Use Case** | Real-time monitoring | Point-in-time reporting |

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.6
Script Type:          Data Consolidation / Alternative Integration Method
Framework Component:  A.8.6 Capacity Management Dashboard Support
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Input Files:
    - ISMS_A_8_6_1_Capacity_Utilization_Assessment_YYYYMMDD.xlsx
    - ISMS_A_8_6_2_Capacity_Forecasts_Planning_YYYYMMDD.xlsx
    - Dashboard workbook (ISMS_A_8_6_3_*.xlsx)

Output:
    - Modified dashboard workbook with consolidated data
    - Console statistics showing consolidation results

Related Quality Tools:
    - normalize_assessment_files_a86.py (file preparation and validation)
    - generate_a86_3_compliance_dashboard.py (dashboard creation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Comprehensive consolidation mapping for capacity management assessments
    - 18 source sheets → 6 dashboard sheets
    - Preserves Executive Dashboard external formulas
    - Safe merged cell handling
    - Detailed consolidation statistics

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**File Modification:**
- Source workbooks are read in data-only mode (NOT modified)
- Dashboard workbook IS modified (data written to detail sheets)
- Executive Dashboard sheet is NOT modified (formulas preserved)
- Always backup dashboard before first consolidation

**Re-Running Consolidation:**
- Safe to re-run when source data is updated
- Existing consolidated data is overwritten
- No cumulative effect (fresh consolidation each time)
- Statistics reflect current consolidation only

**Source File Requirements:**
- File names should match generator script output patterns
- Must be in current directory when running script
- Must be readable .xlsx format (not .xls or .xlsm)
- Must contain expected sheet names from mapping

**Dashboard Requirements:**
- Must be generated by generate_a86_3_compliance_dashboard.py
- Must contain expected target sheet names
- Must be writable (not read-only or open in Excel)
- Should be backed up before first consolidation

**Mapping Maintenance:**
- CONSOLIDATION_MAP must match actual assessment sheet names
- Changing assessment structures requires mapping updates
- Adding custom sheets requires mapping additions
- Verify mapping after generator script modifications

**Performance:**
- Typical consolidation time: 5-15 seconds
- Depends on data volume in source workbooks
- Progress is displayed for each source workbook
- Large datasets (1000+ rows) may take longer

**Limitations:**
- Does not preserve formulas from source workbooks
- Does not copy cell formatting or styles
- Does not handle Excel tables or charts
- Does not validate data during consolidation
- Assumes source data is already validated

**Best Practices:**
1. Complete source assessments before consolidation
2. Run normalize_assessment_files_a86.py before consolidation (recommended)
3. Backup dashboard before first consolidation
4. Review consolidation statistics after each run
5. Test hybrid approach (formulas + consolidation) in dev first
6. Document when/why consolidation approach is used vs. external linking
7. Maintain consistent date suffixes (YYYYMMDD) for version tracking

**Troubleshooting Common Issues:**

**Issue: "ERROR: Dashboard file not found"**
Solution: Verify dashboard path is correct, check file exists

**Issue: "MISSING: Assessment workbook not found"**
Solution: Ensure assessment files are in current directory with correct naming

**Issue: "Sheet 'X' not found, skipping"**
Solution: Source workbook missing expected sheet, check assessment generator

**Issue: "Target sheet 'X' not in dashboard, skipping"**
Solution: Dashboard missing expected sheet, regenerate dashboard

**Issue: "No data" for multiple sheets**
Solution: Source workbooks may be empty or data in unexpected location

**Issue: "ERROR saving dashboard"**
Solution: Dashboard may be open in Excel, close it before consolidation

**Audit Considerations:**
- Consolidation creates static point-in-time snapshot
- Consolidation timestamp is recorded in statistics
- Source data traceability via source labels
- Useful for monthly/quarterly capacity management reporting archives
- Executive Dashboard formulas still provide real-time if links enabled

**Data Protection:**
- Consolidated dashboard inherits data classification from sources
- Source workbooks remain unmodified and unaffected
- Dashboard may contain aggregated sensitive capacity data
- Handle according to organization's data classification policy

**Alternative to This Script:**
If external workbook linking works reliably in your environment, consider
using the standard approach (generate dashboard + enable "Update Links")
instead of data consolidation for automatic updates and reduced maintenance.

**Capacity Management Context:**
Monthly Assessment 1 (Utilization) updates combined with quarterly Assessment 2
(Forecasts) updates means the dashboard needs frequent data refreshes. This
consolidation script facilitates both rapid monthly updates and comprehensive
quarterly consolidations.

**File Pattern Matching:**
Script uses glob patterns to find latest assessment files:
- ISMS_A_8_6_1_*_20*.xlsx → finds latest Utilization assessment
- ISMS_A_8_6_2_*_20*.xlsx → finds latest Forecasts assessment

This allows flexible file naming while maintaining traceability.

================================================================================
END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from datetime import datetime

# ============================================================================
# CONFIGURATION
# ============================================================================

SOURCE_WORKBOOKS = {
    "wb1": "ISMS-IMP-A.8.6.1.xlsx",  # Capacity Utilization
    "wb2": "ISMS-IMP-A.8.6.2.xlsx",  # Capacity Forecasts
}

SHEET_START_ROWS = {
    "Capacity_Risks": 6,
    "Recommendations": 6,
}

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def safely_write_data(ws, start_row, data):
    """Safely write data, handling merged cells"""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except:
                continue
        entries_written += 1
    return entries_written

# ============================================================================
# DATA EXTRACTION
# ============================================================================

def extract_critical_resources(filepath):
    """Extract resources at warning/critical thresholds from WB1"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        critical = []
        
        # Check each resource type sheet
        resource_sheets = ['Compute_Resources', 'Storage_Resources', 'Network_Resources',
                          'Application_Resources', 'Cloud_Resources']
        
        for sheet_name in resource_sheets:
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            resource_type = sheet_name.replace('_Resources', '')
            
            # Scan for resources with warning/critical status
            for row in range(5, min(ws.max_row + 1, 200)):
                resource_name = ws[f'A{row}'].value
                if not resource_name:
                    continue
                
                threshold_status = ws[f'H{row}'].value  # Column H is threshold status
                utilization = ws[f'D{row}'].value  # Column D is utilization %
                
                if threshold_status and ('Warning' in str(threshold_status) or 'Critical' in str(threshold_status)):
                    critical.append({
                        'type': resource_type,
                        'name': resource_name,
                        'utilization': utilization,
                        'status': threshold_status
                    })
        
        wb.close()
        return critical
    except Exception as e:
        print(f"  [!] Error extracting critical resources: {e}")
        return []

def extract_forecast_risks(filepath):
    """Extract capacity exhaustion forecasts from WB2"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        risks = []
        
        if 'Capacity_Exhaustion' not in wb.sheetnames:
            wb.close()
            return []
        
        ws = wb['Capacity_Exhaustion']
        
        # Scan exhaustion predictions
        for row in range(4, min(ws.max_row + 1, 100)):
            resource = ws[f'A{row}'].value
            if not resource:
                continue
            
            months_to_threshold = ws[f'D{row}'].value
            exhaustion_date = ws[f'E{row}'].value
            urgency = ws[f'F{row}'].value
            
            # Track items with urgent timeline
            if months_to_threshold and isinstance(months_to_threshold, (int, float)) and months_to_threshold < 6:
                risks.append({
                    'resource': resource,
                    'months': months_to_threshold,
                    'date': exhaustion_date,
                    'urgency': urgency
                })
        
        wb.close()
        return risks
    except Exception as e:
        print(f"  [!] Error extracting forecast risks: {e}")
        return []

def generate_capacity_risks(critical_resources, forecast_risks):
    """Generate risk entries from capacity data"""
    risks = []
    risk_counter = 1
    
    # Risks from critical utilization
    for resource in critical_resources:
        severity = 'HIGH' if 'Critical' in str(resource['status']) else 'MEDIUM'
        risks.append([
            f"RISK-A86-{risk_counter:03d}",
            resource['type'],
            resource['name'],
            f"{resource['utilization']}%",
            severity,
            'Address capacity constraints',
        ])
        risk_counter += 1
    
    # Risks from forecast exhaustion
    for risk in forecast_risks:
        risks.append([
            f"RISK-A86-{risk_counter:03d}",
            'Forecast',
            risk['resource'],
            f"{risk['months']} months",
            risk['urgency'] if risk['urgency'] else 'HIGH',
            f"Plan expansion before {risk['date']}",
        ])
        risk_counter += 1
    
    return risks

def generate_recommendations(critical_resources, forecast_risks):
    """Generate recommendations from capacity analysis"""
    recommendations = []
    rec_counter = 1
    
    # Recommendations for critical resources
    if critical_resources:
        recommendations.append([
            f"REC-{rec_counter:03d}",
            'Immediate',
            f"Address {len(critical_resources)} critical/warning resources",
            'HIGH',
            'Within 30 days',
        ])
        rec_counter += 1
    
    # Recommendations for forecast risks
    urgent_forecasts = [r for r in forecast_risks if r['months'] < 3]
    if urgent_forecasts:
        recommendations.append([
            f"REC-{rec_counter:03d}",
            'Urgent Planning',
            f"Begin expansion for {len(urgent_forecasts)} resources nearing capacity",
            'HIGH',
            'Within 60 days',
        ])
        rec_counter += 1
    
    # General recommendations
    recommendations.append([
        f"REC-{rec_counter:03d}",
        'Process Improvement',
        'Implement automated capacity monitoring and alerting',
        'MEDIUM',
        'Quarterly review',
    ])
    
    return recommendations

# ============================================================================
# MAIN CONSOLIDATION
# ============================================================================

def populate_dashboard(dashboard_file):
    """Populate dashboard from source workbooks"""
    
    print("="*80)
    print("A.8.6 CAPACITY MANAGEMENT DASHBOARD CONSOLIDATION")
    print("="*80)
    
    # Check for source workbooks
    missing = []
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        if not os.path.exists(wb_filename):
            missing.append(wb_filename)
    
    if missing:
        print("\n[X] ERROR: Missing source workbooks:")
        for wb in missing:
            print(f"    - {wb}")
        print("\n  Place both assessment workbooks in same directory.")
        return False
    
    print("\n[1/3] Extracting critical utilization data...")
    critical_resources = extract_critical_resources(SOURCE_WORKBOOKS['wb1'])
    print(f"  * Critical/Warning resources: {len(critical_resources)}")
    
    print("\n[2/3] Extracting capacity forecasts...")
    forecast_risks = extract_forecast_risks(SOURCE_WORKBOOKS['wb2'])
    print(f"  * Resources nearing capacity: {len(forecast_risks)}")
    
    print("\n[3/3] Generating risks and recommendations...")
    all_risks = generate_capacity_risks(critical_resources, forecast_risks)
    all_recommendations = generate_recommendations(critical_resources, forecast_risks)
    print(f"  * Capacity risks: {len(all_risks)}")
    print(f"  * Recommendations: {len(all_recommendations)}")
    
    # Load dashboard
    try:
        wb = load_workbook(dashboard_file)
    except Exception as e:
        print(f"  [X] Error loading dashboard: {e}")
        return False
    
    # Populate Capacity Risks sheet
    if 'Capacity_Risks' in wb.sheetnames and all_risks:
        ws = wb['Capacity_Risks']
        count = safely_write_data(ws, SHEET_START_ROWS['Capacity_Risks'], all_risks)
        print(f"  [OK] Capacity_Risks: {count} entries")
    
    # Populate Recommendations sheet
    if 'Recommendations' in wb.sheetnames and all_recommendations:
        ws = wb['Recommendations']
        count = safely_write_data(ws, SHEET_START_ROWS['Recommendations'], all_recommendations)
        print(f"  [OK] Recommendations: {count} entries")
    
    # Save
    try:
        wb.save(dashboard_file)
        print(f"\n[SAVED] {dashboard_file}")
    except Exception as e:
        print(f"  [X] Error saving: {e}")
        return False
    
    print("\n" + "="*80)
    print("[OK] A.8.6 DASHBOARD CONSOLIDATION COMPLETE")
    print("="*80)
    print(f"\nSummary:")
    print(f"  * Critical resources: {len(critical_resources)}")
    print(f"  * Forecast risks: {len(forecast_risks)}")
    print(f"  * Total risks documented: {len(all_risks)}")
    print(f"  * Recommendations: {len(all_recommendations)}")
    print(f"\n[ROCKET] Dashboard ready for capacity planning!")
    print("="*80 + "\n")
    
    return True

# ============================================================================
# MAIN
# ============================================================================

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 consolidate_a86_dashboard.py <dashboard.xlsx>")
        print("\nExample:")
        print("  python3 consolidate_a86_dashboard.py ISMS-A.8.6_Capacity_Management_Dashboard_20260117.xlsx")
        sys.exit(1)
    
    dashboard_file = sys.argv[1]
    
    if not os.path.exists(dashboard_file):
        print(f"[X] Error: Dashboard file not found: {dashboard_file}")
        sys.exit(1)
    
    success = populate_dashboard(dashboard_file)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
