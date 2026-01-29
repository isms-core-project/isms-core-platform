#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
Excel Workbook Sanity Checker - ISMS A.8.6.1 Capacity Utilization Assessment
================================================================================

Domain-specific diagnostic utility for A.8.6.1 Capacity Utilization
assessment workbooks.

**Purpose:**
Validates workbook structure, formulas, and data validations specific to
capacity utilization assessment requirements.

**Usage:**
    python3 sanity_check_a86_assessment_1.py ISMS_A_8_6_1_Utilization_YYYYMMDD.xlsx

**Checks Performed:**
- Sheet structure (10 expected sheets)
- Compute capacity resource inventory
- Storage capacity resource inventory
- Network capacity resource inventory
- Application capacity resource inventory
- Cloud resource capacity (if applicable)
- Threshold status validation dropdowns
- Utilization percentage validations (0-100%)
- Capacity headroom calculations
- Evidence register linkage
- Formula integrity in calculated fields
- Conditional formatting for threshold status

**Exit Codes:**
    0 = All checks passed
    1 = Warnings detected (workbook usable)
    2 = Critical errors detected (regenerate recommended)

Control Reference: ISO/IEC 27001:2022 Annex A Control A.8.6
Assessment Type: Assessment 1 of 3 (Current Capacity Utilization)
Version: 1.0
================================================================================
"""

import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def check_a86_1_workbook(filename):
    """Domain-specific validation for A.8.6.1 Capacity Utilization assessment."""
    
    print("=" * 80)
    print("A.8.6.1 CAPACITY UTILIZATION ASSESSMENT - SANITY CHECK")
    print("=" * 80)
    print(f"File: {filename}\n")
    
    issues = []
    warnings = []
    
    # Load workbook
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"✓ Workbook loaded: {len(wb.sheetnames)} sheets\n")
    except Exception as e:
        print(f"✗ CRITICAL: Cannot load workbook: {e}")
        return 2
    
    # ========================================================================
    # EXPECTED SHEET STRUCTURE
    # ========================================================================
    print("=" * 80)
    print("CHECK 1: SHEET STRUCTURE")
    print("=" * 80)
    
    expected_sheets = [
        "Instructions & Legend",
        "Compute_Capacity",
        "Storage_Capacity",
        "Network_Capacity",
        "Application_Capacity",
        "Cloud_Capacity",
        "Capacity_Summary",
        "Threshold_Status",
        "Capacity_Headroom",
        "Peak_Utilization",
        "Gap_Analysis",
        "Evidence_Register",
        "Approval_Sign_Off"
    ]
    
    found_sheets = wb.sheetnames
    missing_sheets = [s for s in expected_sheets if s not in found_sheets]
    extra_sheets = [s for s in found_sheets if s not in expected_sheets]
    
    if missing_sheets:
        for sheet in missing_sheets:
            issues.append(f"  ✗ Missing required sheet: '{sheet}'")
    
    if extra_sheets:
        for sheet in extra_sheets:
            warnings.append(f"  ⚠ Unexpected sheet found: '{sheet}'")
    
    if not missing_sheets and not extra_sheets:
        print("  ✓ All expected sheets present, no extras")
    
    print(f"  Expected: {len(expected_sheets)} sheets")
    print(f"  Found: {len(found_sheets)} sheets")
    
    # ========================================================================
    # COMPUTE CAPACITY VALIDATION
    # ========================================================================
    if "Compute_Capacity" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 2: COMPUTE CAPACITY DATA VALIDATION")
        print("=" * 80)
        
        ws = wb["Compute_Capacity"]
        
        expected_validations = {
            'Resource_Type': ['Physical Server', 'Virtual Machine', 'Container', 'Serverless'],
            'Environment': ['Production', 'Staging', 'Development', 'Testing', 'DR'],
            'Threshold_Status': ['OK', 'Warning', 'Critical'],
            'Monitoring_Status': ['Active', 'Inactive', 'Partial']
        }
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
            
            validation_found = {key: False for key in expected_validations.keys()}
            
            for dv in ws.data_validations.dataValidation:
                if dv.formula1:
                    formula = str(dv.formula1)
                    for key, values in expected_validations.items():
                        if any(v in formula for v in values):
                            validation_found[key] = True
            
            for key, found in validation_found.items():
                if found:
                    print(f"  ✓ {key} validation detected")
                else:
                    warnings.append(f"  ⚠ {key} validation not detected")
        else:
            warnings.append("  ⚠ No data validations found in Compute Capacity")
    
    # ========================================================================
    # STORAGE CAPACITY VALIDATION
    # ========================================================================
    if "Storage_Capacity" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 3: STORAGE CAPACITY DATA VALIDATION")
        print("=" * 80)
        
        ws = wb["Storage_Capacity"]
        
        expected_storage_types = ['SAN', 'NAS', 'Object Storage', 'Block Storage', 
                                  'Database', 'Backup', 'Archive']
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
            
            storage_type_found = False
            for dv in ws.data_validations.dataValidation:
                if dv.formula1:
                    formula = str(dv.formula1)
                    if any(st in formula for st in expected_storage_types):
                        storage_type_found = True
                        print("  ✓ Storage type validation detected")
                        break
            
            if not storage_type_found:
                warnings.append("  ⚠ Storage type validation not detected")
        else:
            warnings.append("  ⚠ No data validations in Storage Capacity")
    
    # ========================================================================
    # NETWORK CAPACITY VALIDATION
    # ========================================================================
    if "Network_Capacity" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 4: NETWORK CAPACITY DATA VALIDATION")
        print("=" * 80)
        
        ws = wb["Network_Capacity"]
        
        expected_network_types = ['WAN Link', 'Internet Connection', 'LAN Switch', 
                                  'Load Balancer', 'Firewall', 'VPN Gateway']
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
            
            network_type_found = False
            for dv in ws.data_validations.dataValidation:
                if dv.formula1:
                    formula = str(dv.formula1)
                    if any(nt in formula for nt in expected_network_types):
                        network_type_found = True
                        print("  ✓ Network type validation detected")
                        break
            
            if not network_type_found:
                warnings.append("  ⚠ Network type validation not detected")
        else:
            warnings.append("  ⚠ No data validations in Network Capacity")
    
    # ========================================================================
    # APPLICATION CAPACITY VALIDATION
    # ========================================================================
    if "Application_Capacity" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 5: APPLICATION CAPACITY DATA VALIDATION")
        print("=" * 80)
        
        ws = wb["Application_Capacity"]
        
        expected_metrics = ['Concurrent Users', 'Transactions/sec', 'Queue Depth', 
                           'API Rate Limit', 'Session Count']
        
        metrics_found = []
        for row in ws.iter_rows(min_row=1, max_row=20, max_col=15):
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value)
                    for metric in expected_metrics:
                        if metric in cell_value and metric not in metrics_found:
                            metrics_found.append(metric)
        
        print(f"  Application metrics detected: {len(metrics_found)}/{len(expected_metrics)}")
        for metric in expected_metrics:
            if metric in metrics_found:
                print(f"  ✓ {metric} metric detected")
        
        if len(metrics_found) < 2:
            warnings.append("  ⚠ Limited application capacity metrics detected")
    
    # ========================================================================
    # THRESHOLD STATUS VALIDATION
    # ========================================================================
    if "Threshold_Status" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 6: THRESHOLD STATUS SUMMARY")
        print("=" * 80)
        
        ws = wb["Threshold_Status"]
        
        expected_thresholds = ['OK', 'Warning', 'Critical']
        
        thresholds_found = []
        for row in ws.iter_rows(min_row=1, max_row=30, max_col=10):
            for cell in row:
                if cell.value and str(cell.value) in expected_thresholds:
                    if str(cell.value) not in thresholds_found:
                        thresholds_found.append(str(cell.value))
        
        print(f"  Threshold statuses detected: {len(thresholds_found)}/{len(expected_thresholds)}")
        for threshold in expected_thresholds:
            if threshold in thresholds_found:
                print(f"  ✓ {threshold} threshold status found")
            else:
                warnings.append(f"  ⚠ {threshold} threshold not found in data")
    
    # ========================================================================
    # CAPACITY HEADROOM VALIDATION
    # ========================================================================
    if "Capacity_Headroom" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 7: CAPACITY HEADROOM CALCULATIONS")
        print("=" * 80)
        
        ws = wb["Capacity_Headroom"]
        
        # Check for formulas (headroom should be calculated)
        formula_count = 0
        for row in ws.iter_rows(min_row=5, max_row=100):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
        
        if formula_count > 0:
            print(f"  ✓ Formulas detected: {formula_count}")
            print("  ✓ Capacity headroom calculations present")
        else:
            warnings.append("  ⚠ No calculation formulas detected in Capacity Headroom")
    
    # ========================================================================
    # FORMULA INTEGRITY CHECK
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 8: FORMULA INTEGRITY")
    print("=" * 80)
    
    formula_count = 0
    formula_errors = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
                    formula = cell.value
                    
                    # Check for common errors
                    if formula.count('(') != formula.count(')'):
                        issues.append(f"  ✗ {sheet_name}!{cell.coordinate}: Unbalanced parentheses")
                        formula_errors += 1
                    
                    if formula.count('"') % 2 != 0:
                        issues.append(f"  ✗ {sheet_name}!{cell.coordinate}: Unbalanced quotes")
                        formula_errors += 1
                    
                    # Check for invalid percentage calculations
                    if '%' in formula and 'TEXT' not in formula:
                        # Warn about potential percentage calculation issues
                        pass  # This is actually fine, just noting
    
    print(f"  Total formulas found: {formula_count}")
    if formula_errors == 0:
        print("  ✓ No formula syntax errors detected")
    else:
        print(f"  ✗ Formula errors found: {formula_errors}")
    
    # ========================================================================
    # UTILIZATION PERCENTAGE VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 9: UTILIZATION PERCENTAGE VALIDATION")
    print("=" * 80)
    
    # Check for percentage values in key capacity sheets
    capacity_sheets = ["Compute_Capacity", "Storage_Capacity", "Network_Capacity", 
                       "Application_Capacity"]
    
    invalid_percentages = []
    
    for sheet_name in capacity_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=10, max_row=100):
                for cell in row:
                    if cell.value and isinstance(cell.value, (int, float)):
                        if cell.number_format and '%' in cell.number_format:
                            # Cell is formatted as percentage
                            if cell.value > 1.0:  # Excel stores 100% as 1.0
                                invalid_percentages.append(f"{sheet_name}!{cell.coordinate}: {cell.value*100:.1f}%")
                            elif cell.value < 0:
                                invalid_percentages.append(f"{sheet_name}!{cell.coordinate}: {cell.value*100:.1f}%")
    
    if invalid_percentages:
        print(f"  ⚠ Invalid percentage values found: {len(invalid_percentages)}")
        for inv_pct in invalid_percentages[:5]:  # Show first 5
            warnings.append(f"  ⚠ {inv_pct}")
        if len(invalid_percentages) > 5:
            warnings.append(f"  ⚠ ... and {len(invalid_percentages)-5} more")
    else:
        print("  ✓ All percentage values within valid range (0-100%)")
    
    # ========================================================================
    # CONDITIONAL FORMATTING CHECK
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 10: CONDITIONAL FORMATTING")
    print("=" * 80)
    
    total_cf_rules = 0
    sheets_with_cf = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
            cf_count = len(ws.conditional_formatting._cf_rules)
            if cf_count > 0:
                sheets_with_cf += 1
                total_cf_rules += cf_count
    
    print(f"  Total conditional formatting rules: {total_cf_rules}")
    print(f"  Sheets with conditional formatting: {sheets_with_cf}")
    
    if total_cf_rules > 0:
        print("  ✓ Conditional formatting present (threshold status colors)")
    else:
        warnings.append("  ⚠ No conditional formatting detected")
    
    # ========================================================================
    # EVIDENCE REGISTER LINKAGE
    # ========================================================================
    if "Evidence_Register" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 11: EVIDENCE REGISTER STRUCTURE")
        print("=" * 80)
        
        ws = wb["Evidence_Register"]
        
        has_content = False
        for row in ws.iter_rows(min_row=1, max_row=15, max_col=10):
            for cell in row:
                if cell.value:
                    has_content = True
                    break
            if has_content:
                break
        
        if has_content:
            print("  ✓ Evidence Register has content")
        else:
            warnings.append("  ⚠ Evidence Register appears empty")
        
        # Check for data validations
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
        else:
            warnings.append("  ⚠ No data validations in Evidence Register")
    
    # ========================================================================
    # GAP ANALYSIS STRUCTURE
    # ========================================================================
    if "Gap_Analysis" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 12: GAP ANALYSIS STRUCTURE")
        print("=" * 80)
        
        ws = wb["Gap_Analysis"]
        
        expected_severities = ['Critical', 'High', 'Medium', 'Low']
        expected_statuses = ['Open', 'In Progress', 'Planned', 'Resolved']
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
            
            severity_found = False
            status_found = False
            
            for dv in ws.data_validations.dataValidation:
                if dv.formula1:
                    formula = str(dv.formula1)
                    if any(s in formula for s in expected_severities):
                        severity_found = True
                    if any(s in formula for s in expected_statuses):
                        status_found = True
            
            if severity_found:
                print("  ✓ Severity validation detected")
            else:
                warnings.append("  ⚠ Severity validation not detected")
            
            if status_found:
                print("  ✓ Status validation detected")
            else:
                warnings.append("  ⚠ Status validation not detected")
        else:
            warnings.append("  ⚠ No data validations in Gap Analysis")
    
    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    print("\n" + "=" * 80)
    print("VALIDATION SUMMARY")
    print("=" * 80)
    
    if issues:
        print(f"\n❌ CRITICAL ISSUES: {len(issues)}")
        for issue in issues:
            print(issue)
    
    if warnings:
        print(f"\n⚠️  WARNINGS: {len(warnings)}")
        for warning in warnings:
            print(warning)
    
    if not issues and not warnings:
        print("\n✅ ALL CHECKS PASSED")
        print("\nThe A.8.6.1 Capacity Utilization assessment workbook appears healthy.")
        print("Ready for data collection and capacity analysis.")
    else:
        print("\n" + "=" * 80)
        print("RECOMMENDATIONS:")
        print("=" * 80)
        
        if missing_sheets:
            print("\n• Regenerate workbook with generate_a86_1_capacity_utilization.py")
            print("  to ensure all required sheets are present")
        
        if invalid_percentages:
            print("\n• Review utilization percentages:")
            print("  - Valid range: 0-100%")
            print("  - Values >100% indicate data quality issues")
            print("  - Negative values indicate calculation errors")
        
        if warnings:
            print("\n• Review warnings - most are informational")
            print("• Data validations improve data quality but are optional")
            print("• Ensure capacity monitoring data is current and accurate")
    
    print("\n" + "=" * 80)
    
    # Return exit code
    if issues:
        return 2
    elif warnings:
        return 1
    else:
        return 0


def main():
    if len(sys.argv) < 2:
        print("=" * 80)
        print("A.8.6.1 CAPACITY UTILIZATION ASSESSMENT - SANITY CHECK")
        print("=" * 80)
        print("\nUsage: python3 sanity_check_a86_assessment_1.py <filename.xlsx>")
        print("\nExample:")
        print("  python3 sanity_check_a86_assessment_1.py ISMS_A_8_6_1_Utilization_20250128.xlsx")
        print("\nExit codes:")
        print("  0 = All checks passed")
        print("  1 = Warnings detected (workbook usable)")
        print("  2 = Critical errors detected (regenerate recommended)")
        print("\n" + "=" * 80)
        sys.exit(1)
    
    filename = sys.argv[1]
    exit_code = check_a86_1_workbook(filename)
    sys.exit(exit_code)


if __name__ == "__main__":
    main()