#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.6-Assessment-1 - Capacity Utilization Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.6: Capacity Management
Assessment Workbook 1 of 3: Current Capacity Utilization Analysis

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific infrastructure, monitoring systems, and capacity
management requirements.

Key customization areas:
1. Resource inventory (servers, storage, networks - match your infrastructure)
2. Utilization thresholds (warning/critical levels - adapt to your risk profile)
3. Monitoring tool integrations (Prometheus, Datadog, CloudWatch, etc.)
4. Capacity headroom calculations (based on your planning methodology)
5. Assessment frequency and reporting cycles (monthly, quarterly)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.6 Capacity Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for analyzing
current capacity utilization across all infrastructure resources to support
proactive capacity planning and prevent resource exhaustion.

**Purpose:**
Enables systematic assessment of current resource utilization against ISO
27001:2022 Control A.8.6 requirements, supporting evidence-based capacity
planning decisions and preventing capacity-related service degradation.

**Assessment Scope:**
- Compute capacity utilization (CPU, memory, VM/container capacity)
- Storage capacity utilization (disk space, database storage, backup storage)
- Network capacity utilization (bandwidth, throughput, load balancer capacity)
- Application capacity utilization (concurrent users, transaction rates, queues)
- Cloud service capacity (cloud quotas, service limits, cost thresholds)
- Threshold status evaluation (OK, Warning, Critical)
- Capacity headroom calculation (remaining capacity before exhaustion)
- Peak utilization tracking (maximum observed usage)
- Utilization trends (growth patterns over time)
- Gap identification for resources exceeding thresholds

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and capacity standards
2. Compute Capacity - Server CPU, memory, virtualization capacity
3. Storage Capacity - Disk space, database, backup, archive storage
4. Network Capacity - Bandwidth, throughput, network infrastructure
5. Application Capacity - User sessions, transactions, queue depths
6. Cloud Capacity - Cloud service quotas and limits (if applicable)
7. Capacity Summary - Consolidated view of all resource utilization
8. Threshold Status - Resources by threshold status (OK/Warning/Critical)
9. Capacity Headroom - Available capacity before resource exhaustion
10. Peak Utilization - Maximum observed capacity usage
11. Gap Analysis - Resources requiring immediate capacity expansion
12. Evidence Register - Supporting documentation and monitoring data
13. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with threshold status dropdown lists
- Conditional formatting for utilization status visualization
- Automated headroom calculations (remaining capacity percentage)
- Protected formulas with unprotected input cells
- Peak utilization tracking for capacity planning
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with capacity forecasting assessment (Assessment 2)

**Integration:**
This assessment provides current utilization data for the A.8.6 Capacity
Management Dashboard, which consolidates utilization, forecasts, and planning
effectiveness for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a86_1_capacity_utilization.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a86_1_capacity_utilization.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a86_1_capacity_utilization.py --date 20250128

Output:
    File: ISMS_A_8_6_1_Capacity_Utilization_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize capacity thresholds to match your risk profile
    2. Inventory all infrastructure resources requiring capacity monitoring
    3. Export current utilization data from monitoring tools
    4. Complete utilization assessments for each resource type
    5. Calculate capacity headroom (remaining capacity)
    6. Identify resources at warning/critical thresholds
    7. Review peak utilization patterns for trend analysis
    8. Conduct gap analysis for resources requiring expansion
    9. Collect and link evidence (monitoring dashboards, reports)
    10. Obtain stakeholder approvals
    11. Feed results into A.8.6 Capacity Management Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.6
Assessment Type:      Assessment 1 of 3 (Current Capacity Utilization)
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.6: Capacity Management Policy (Governance)
    - ISMS-IMP-A.8.6-S1: Capacity Monitoring Implementation Guide
    - ISMS-IMP-A.8.6-S2: Capacity Forecasting & Planning Implementation Guide
    - ISMS-IMP-A.8.6-S3: Capacity Management Assessment Guide
    - Assessment 2: Capacity Forecasts & Planning (generate_a86_2_capacity_forecasts.py)
    - Dashboard: Capacity Management Overview (generate_a86_3_compliance_dashboard.py)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.6-S3 specification
    - Supports comprehensive capacity utilization evaluation
    - Integrated with A.8.6 Capacity Management Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Capacity Thresholds:**
Capacity thresholds vary by resource type and organizational risk tolerance.
Default thresholds (Warning: 70%, Critical: 85%) should be reviewed and
customized based on:
- Resource type criticality (production vs. development)
- Lead time for capacity expansion (hardware procurement vs. cloud scaling)
- Performance degradation characteristics (gradual vs. sudden)
- Business impact tolerance (customer-facing vs. internal systems)

**Monitoring Tool Integration:**
This assessment workbook is tool-agnostic and accepts data from any monitoring
platform including:
- Open source: Prometheus, Grafana, Nagios, Zabbix
- Commercial SaaS: Datadog, New Relic, Dynatrace, Splunk
- Cloud-native: AWS CloudWatch, Azure Monitor, GCP Cloud Monitoring
- Enterprise: IBM Tivoli, HP OpenView, BMC TrueSight

Export utilization data from your monitoring tool and populate the assessment
workbook manually or via CSV import (see Implementation Guide).

**Assessment Frequency:**
Monthly assessment is recommended for most environments:
- Monthly: Tactical capacity review and immediate action items
- Quarterly: Comprehensive capacity planning (feeds Assessment 2)
- Annual: Strategic capacity planning and budget development

Critical production systems may require weekly or daily capacity reviews.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Current utilization data from monitoring systems
- Documentation of threshold definitions and rationale
- Evidence of capacity planning actions for resources at warning/critical
- Approval from infrastructure management and IT leadership

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- System inventory and resource specifications
- Capacity utilization patterns and performance data
- Capacity constraints and bottleneck identification
- Infrastructure growth trends and projections

Handle in accordance with your organization's data classification policies.
Consider this assessment workbook "Internal" or "Confidential" classification.

**Maintenance:**
Review and update assessment:
- Monthly: Update utilization data and threshold status
- Quarterly: Review capacity headroom and expansion requirements
- Annually: Validate threshold definitions and assessment methodology
- Ad-hoc: When infrastructure changes or new resources are deployed

**Quality Assurance:**
Have infrastructure managers and capacity planning team validate assessments
before using results for capacity planning decisions or compliance reporting.
Cross-reference utilization data against monitoring tool dashboards to ensure
accuracy.

**Integration with Forecasting:**
Assessment 1 (Utilization) provides current-state data that feeds into
Assessment 2 (Forecasts & Planning) for trend analysis and future capacity
projection. Ensure both assessments are completed for comprehensive capacity
management.

**Cloud vs. On-Premises Considerations:**
- **On-premises**: Focus on hardware capacity limits and procurement lead times
- **Cloud**: Focus on service quotas, auto-scaling policies, and cost thresholds
- **Hybrid**: Assess both on-premises and cloud capacity independently

Customize the assessment to match your infrastructure deployment model.

================================================================================
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation




# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅
XMARK = '\u274C'      # ❌
WARNING = '\u26A0'    # ⚠️
HOURGLASS = '\u23F3'  # ⏳
BULLET = '\u2022'     # •
ARROW = '\u2192'      # →
CHART = '\U0001F4CA' # 📊
TARGET = '\U0001F3AF' # 🎯
TRENDING_UP = '\U0001F4C8' # 📈
DISK = '\U0001F4BE'   # 💾
STOPWATCH = '\u23F1'  # ⏱️

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches ISMS-IMP-A.8.6.1 specification
    sheets = [
        "Instructions & Legend",
        "Compute_Resources",
        "Storage_Resources",
        "Network_Resources",
        "Application_Resources",
        "Cloud_Resources",
        "Threshold_Summary",
        "Coverage_Analysis",
        "Evidence_Register",
        "Approval_Sign_Of",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_ok": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_warning": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_critical": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
    }
    return styles


def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & COLUMN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    Returns dictionary of validation objects that can be applied to cells.
    """
    validations = {}
    
    # Status validation
    validations['status'] = DataValidation(
        type="list",
        formula1='f"{CHECK} OK,⚠️ Warning,❌ Critical,Not Monitored"',
        allow_blank=False
    )
    
    # Yes/No validation
    validations['yes_no'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    
    # Resource type validation
    validations['resource_type'] = DataValidation(
        type="list",
        formula1='"Physical Server,Virtual Machine,Container/Pod,Cloud Instance,Other"',
        allow_blank=False
    )
    
    # Monitoring tool validation
    validations['monitoring_tool'] = DataValidation(
        type="list",
        formula1='"Prometheus,Datadog,New Relic,CloudWatch,Azure Monitor,GCP Monitoring,Zabbix,Nagios,PRTG,Other"',
        allow_blank=True
    )
    
    # Add all validations to worksheet
    for validation in validations.values():
        ws.add_data_validation(validation)
    
    return validations


# ============================================================================
# SECTION 3: SHEET 1 - INSTRUCTIONS & LEGEND
# ============================================================================

def create_instructions(wb, styles):
    """Create Instructions & Legend sheet."""
    ws = wb["Instructions & Legend"]
    
    # Header
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "ISMS-IMP-A.8.6.1 – Capacity Utilization Assessment"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = "ISO/IEC 27001:2022 - Control A.8.6: Capacity Management"
    apply_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 25
    
    # Document information
    row = 4
    info = [
        ("Document ID:", "ISMS-IMP-A.8.6.1"),
        ("Assessment Area:", "Capacity Utilization Analysis"),
        ("Related Policy:", "ISMS-POL-A.8.6-S2 (Capacity Management Policy)"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Monthly"),
    ]
    
    for label, value in info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if "USER INPUT" in value:
            apply_style(ws[f'B{row}'], styles['input_cell'])
        row += 1
    
    # How to use
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    cell.value = "HOW TO USE THIS WORKBOOK"
    apply_style(cell, styles['subheader'])
    
    row += 1
    instructions = [
        "1. Complete resource inventory in each resource type sheet (Compute, Storage, Network, Application, Cloud)",
        "2. Enter current utilization data from monitoring tools",
        "3. Utilization percentages and threshold status will auto-calculate",
        "4. Review Threshold_Summary for overview of capacity health",
        "5. Complete Coverage_Analysis to identify monitoring gaps",
        "6. Document evidence in Evidence_Register",
        "7. Obtain final approval and sign-of",
        "",
        "Yellow cells = User input required",
        "White cells = Auto-calculated (do not edit)",
    ]
    
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        row += 1
    
    # Status legend
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    cell.value = "STATUS LEGEND"
    apply_style(cell, styles['subheader'])
    
    row += 1
    ws[f'A{row}'] = "Symbol"
    ws[f'B{row}'] = "Status"
    ws[f'C{row}'] = "Description"
    ws[f'D{row}'] = "Typical Threshold"
    apply_style(ws[f'A{row}'], styles['column_header'])
    apply_style(ws[f'B{row}'], styles['column_header'])
    apply_style(ws[f'C{row}'], styles['column_header'])
    apply_style(ws[f'D{row}'], styles['column_header'])
    
    row += 1
    legend = [
        (f"{CHECK}", "OK", "Utilization below warning threshold", "< 70%"),
        (f"{WARNING}", "Warning", "Capacity planning should begin", "70-85%"),
        (f"{XMARK}", "Critical", "Immediate action required", "> 85%"),
        ("Not Monitored", "Not Monitored", "Resource not yet monitored", "N/A"),
    ]
    
    for symbol, status, desc, threshold in legend:
        ws[f'A{row}'] = symbol
        ws[f'B{row}'] = status
        ws[f'C{row}'] = desc
        ws[f'D{row}'] = threshold
        if status == "OK":
            apply_style(ws[f'B{row}'], styles['status_ok'])
        elif status == "Warning":
            apply_style(ws[f'B{row}'], styles['status_warning'])
        elif status == "Critical":
            apply_style(ws[f'B{row}'], styles['status_critical'])
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 20


# ============================================================================
# SECTION 4: SHEET 2 - COMPUTE RESOURCES
# ============================================================================

def create_compute_resources(wb, styles):
    """Create Compute Resources capacity sheet."""
    ws = wb["Compute_Resources"]
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:M1')
    cell = ws['A1']
    cell.value = "COMPUTE CAPACITY UTILIZATION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:M2')
    cell = ws['A2']
    cell.value = "Document current CPU and memory capacity utilization"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Resource Name", "Resource Type", "Location/Cluster", "Business Function",
        "Total CPU (Cores)", "Used CPU (Cores)", "CPU Utilization (%)",
        "Total Memory (GB)", "Used Memory (GB)", "Memory Utilization (%)",
        "Peak CPU (%)", "Peak Memory (%)", "Threshold Status"
    ]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Example data row
    row = 5
    example = [
        "server01.example.com", "Physical Server", "DC1-Rack5", "Web Application",
        "16", "8", "", "64", "32", "", "75", "60", ""
    ]
    
    for col, value in enumerate(example, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        if col in [1, 2, 3, 4, 5, 8, 9, 11, 12]:  # Input cells
            apply_style(cell, styles['input_cell'])
    
    # Add formulas for utilization percentage and status
    ws['G5'] = '=IF(AND(E5<>"", F5<>""), ROUND((F5/E5)*100, 1), "")'
    ws['J5'] = '=IF(AND(H5<>"", I5<>""), ROUND((I5/H5)*100, 1), "")'
    ws['M5'] = '=IF(OR(G5="", J5=""), "Not Monitored", IF(OR(G5>85, J5>90), f"{XMARK} Critical", IF(OR(G5>70, J5>75), f"{WARNING} Warning", f"{CHECK} OK")))'
    
    # Apply validations
    validations['resource_type'].add(f'B5:B100')
    validations['status'].add(f'M5:M100')
    
    # Column widths
    widths = [30, 20, 20, 25, 15, 15, 18, 15, 15, 20, 15, 15, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ============================================================================
# SECTION 5: SHEET 3 - STORAGE RESOURCES
# ============================================================================

def create_storage_resources(wb, styles):
    """Create Storage Resources capacity sheet."""
    ws = wb["Storage_Resources"]
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "STORAGE CAPACITY UTILIZATION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:K2')
    cell = ws['A2']
    cell.value = "Document current disk, database, and backup storage capacity"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Resource Name", "Storage Type", "Location/Server", "Business Function",
        "Total Capacity (GB)", "Used Capacity (GB)", "Available Capacity (GB)",
        "Utilization (%)", "Peak Utilization (%)", "Growth Rate (GB/month)", "Threshold Status"
    ]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Example data row
    row = 5
    example = [
        "/dev/sda1 (/)", "Filesystem", "server01", "Operating System",
        "500", "375", "", "", "80", "5", ""
    ]
    
    for col, value in enumerate(example, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        if col in [1, 2, 3, 4, 5, 6, 9, 10]:  # Input cells
            apply_style(cell, styles['input_cell'])
    
    # Add formulas
    ws['G5'] = '=IF(AND(E5<>"", F5<>""), E5-F5, "")'
    ws['H5'] = '=IF(AND(E5<>"", F5<>""), ROUND((F5/E5)*100, 1), "")'
    ws['K5'] = '=IF(H5="", "Not Monitored", IF(H5>85, f"{XMARK} Critical", IF(H5>75, f"{WARNING} Warning", f"{CHECK} OK")))'
    
    # Apply storage type validation
    storage_types = DataValidation(
        type="list",
        formula1='"Filesystem,Database,Backup Repository,Archive,SAN/NAS,Object Storage,Other"',
        allow_blank=False
    )
    ws.add_data_validation(storage_types)
    storage_types.add('B5:B100')
    
    validations['status'].add(f'K5:K100')
    
    # Column widths
    widths = [30, 20, 25, 25, 18, 18, 20, 15, 18, 20, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ============================================================================
# SECTION 6: SHEET 4 - NETWORK RESOURCES
# ============================================================================

def create_network_resources(wb, styles):
    """Create Network Resources capacity sheet."""
    ws = wb["Network_Resources"]
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:L1')
    cell = ws['A1']
    cell.value = "NETWORK CAPACITY UTILIZATION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:L2')
    cell = ws['A2']
    cell.value = "Document network bandwidth, throughput, and service capacity"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Resource Name", "Network Type", "Location", "Business Function",
        "Link Capacity (Mbps)", "Avg Inbound (Mbps)", "Avg Outbound (Mbps)",
        "Total Utilization (%)", "Peak Utilization (%)", "Concurrent Connections",
        "Connection Limit", "Threshold Status"
    ]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Example data row
    row = 5
    example = [
        "WAN-Link-Primary", "WAN/Internet", "DC1", "Primary Internet",
        "1000", "350", "150", "", "75", "5000", "10000", ""
    ]
    
    for col, value in enumerate(example, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        if col in [1, 2, 3, 4, 5, 6, 7, 9, 10, 11]:  # Input cells
            apply_style(cell, styles['input_cell'])
    
    # Add formulas
    ws['H5'] = '=IF(AND(E5<>"", F5<>"", G5<>""), ROUND(((F5+G5)/E5)*100, 1), "")'
    ws['L5'] = '=IF(H5="", "Not Monitored", IF(H5>85, f"{XMARK} Critical", IF(H5>70, f"{WARNING} Warning", f"{CHECK} OK")))'
    
    # Network type validation
    network_types = DataValidation(
        type="list",
        formula1='"WAN/Internet,LAN,Inter-DC Link,VPN,Load Balancer,Firewall,Other"',
        allow_blank=False
    )
    ws.add_data_validation(network_types)
    network_types.add('B5:B100')
    
    validations['status'].add(f'L5:L100')
    
    # Column widths
    widths = [25, 20, 20, 25, 18, 18, 18, 18, 18, 20, 15, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ============================================================================
# SECTION 7: SHEET 5 - APPLICATION RESOURCES
# ============================================================================

def create_application_resources(wb, styles):
    """Create Application Resources capacity sheet."""
    ws = wb["Application_Resources"]
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "APPLICATION CAPACITY UTILIZATION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:K2')
    cell = ws['A2']
    cell.value = "Document application user sessions, transactions, and performance capacity"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Application Name", "Application Type", "Business Function",
        "Max Concurrent Users", "Current Active Users", "User Utilization (%)",
        "Max Transactions/sec", "Current Transactions/sec", "Transaction Utilization (%)",
        "Avg Response Time (ms)", "Threshold Status"
    ]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Example data row
    row = 5
    example = [
        "WebApp-Production", "Web Application", "Customer Portal",
        "1000", "650", "", "500", "280", "", "145", ""
    ]
    
    for col, value in enumerate(example, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        if col in [1, 2, 3, 4, 5, 7, 8, 10]:  # Input cells
            apply_style(cell, styles['input_cell'])
    
    # Add formulas
    ws['F5'] = '=IF(AND(D5<>"", E5<>""), ROUND((E5/D5)*100, 1), "")'
    ws['I5'] = '=IF(AND(G5<>"", H5<>""), ROUND((H5/G5)*100, 1), "")'
    ws['K5'] = '=IF(OR(F5="", I5=""), "Not Monitored", IF(OR(F5>90, I5>90), f"{XMARK} Critical", IF(OR(F5>75, I5>75), f"{WARNING} Warning", f"{CHECK} OK")))'
    
    # Application type validation
    app_types = DataValidation(
        type="list",
        formula1='"Web Application,Database,API Service,Message Queue,Batch Processing,SaaS Application,Other"',
        allow_blank=False
    )
    ws.add_data_validation(app_types)
    app_types.add('B5:B100')
    
    validations['status'].add(f'K5:K100')
    
    # Column widths
    widths = [25, 20, 25, 20, 20, 18, 20, 22, 22, 22, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ============================================================================
# SECTION 8: SHEET 6 - CLOUD RESOURCES
# ============================================================================

def create_cloud_resources(wb, styles):
    """Create Cloud Resources capacity sheet."""
    ws = wb["Cloud_Resources"]
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "CLOUD SERVICE CAPACITY UTILIZATION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "Document cloud service quotas, limits, and instance counts"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    headers = [
        "Cloud Provider", "Service/Resource Type", "Region",
        "Quota/Limit", "Current Usage", "Utilization (%)",
        "Monthly Cost ($)", "Reserved Capacity", "Auto-Scaling Enabled", "Threshold Status"
    ]
    
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Example data row
    row = 5
    example = [
        "AWS", "EC2 Instances (t3.large)", "us-east-1",
        "100", "72", "", "4320", "20 Reserved", "Yes", ""
    ]
    
    for col, value in enumerate(example, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        if col in [1, 2, 3, 4, 5, 7, 8, 9]:  # Input cells
            apply_style(cell, styles['input_cell'])
    
    # Add formulas
    ws['F5'] = '=IF(AND(D5<>"", E5<>""), ROUND((E5/D5)*100, 1), "")'
    ws['J5'] = '=IF(F5="", "Not Monitored", IF(F5>85, f"{XMARK} Critical", IF(F5>70, f"{WARNING} Warning", f"{CHECK} OK")))'
    
    # Cloud provider validation
    cloud_providers = DataValidation(
        type="list",
        formula1='"AWS,Azure,GCP,Oracle Cloud,IBM Cloud,Alibaba Cloud,Other"',
        allow_blank=False
    )
    ws.add_data_validation(cloud_providers)
    cloud_providers.add('A5:A100')
    
    validations['yes_no'].add('I5:I100')
    validations['status'].add('J5:J100')
    
    # Column widths
    widths = [15, 30, 15, 15, 15, 15, 18, 20, 20, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ============================================================================
# SECTION 9: SHEET 7 - THRESHOLD SUMMARY
# ============================================================================

def create_threshold_summary(wb, styles):
    """Create Threshold Summary dashboard sheet."""
    ws = wb["Threshold_Summary"]
    
    # Header
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "CAPACITY THRESHOLD STATUS SUMMARY"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "Aggregate view of resources by threshold status"
    apply_style(cell, styles['subheader'])
    
    # Summary metrics
    row = 4
    ws[f'A{row}'] = "Resource Category"
    ws[f'B{row}'] = "Total Resources"
    ws[f'C{row}'] = f"{CHECK} OK"
    ws[f'D{row}'] = f"{WARNING} Warning"
    ws[f'E{row}'] = f"{XMARK} Critical"
    ws[f'F{row}'] = "Not Monitored"
    
    for col in range(1, 7):
        apply_style(ws.cell(row=row, column=col), styles['column_header'])
    
    # Resource categories with formulas
    categories = [
        ("Compute Resources", "Compute_Resources", "M"),
        ("Storage Resources", "Storage_Resources", "K"),
        ("Network Resources", "Network_Resources", "L"),
        ("Application Resources", "Application_Resources", "K"),
        ("Cloud Resources", "Cloud_Resources", "J"),
    ]
    
    row = 5
    for category, sheet, col in categories:
        ws[f'A{row}'] = category
        ws[f'B{row}'] = f'=COUNTA({sheet}!{col}5:{col}100)'
        ws[f'C{row}'] = f'=COUNTIF({sheet}!{col}5:{col}100,f"{CHECK} OK")'
        ws[f'D{row}'] = f'=COUNTIF({sheet}!{col}5:{col}100,f"{WARNING} Warning")'
        ws[f'E{row}'] = f'=COUNTIF({sheet}!{col}5:{col}100,f"{XMARK} Critical")'
        ws[f'F{row}'] = f'=COUNTIF({sheet}!{col}5:{col}100,"Not Monitored")'
        
        apply_style(ws[f'C{row}'], styles['status_ok'])
        apply_style(ws[f'D{row}'], styles['status_warning'])
        apply_style(ws[f'E{row}'], styles['status_critical'])
        row += 1
    
    # Total row
    row += 1
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f'=SUM(B5:B{row-1})'
    ws[f'C{row}'] = f'=SUM(C5:C{row-1})'
    ws[f'D{row}'] = f'=SUM(D5:D{row-1})'
    ws[f'E{row}'] = f'=SUM(E5:E{row-1})'
    ws[f'F{row}'] = f'=SUM(F5:F{row-1})'
    
    for col in range(2, 7):
        ws.cell(row=row, column=col).font = Font(bold=True)
    
    # Capacity health score
    row += 2
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "CAPACITY HEALTH SCORE"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = f'=IF(B{row-2}>0, ROUND((C{row-2}/B{row-2})*100, 1), 0)'
    ws[f'D{row}'] = "%"
    ws[f'C{row}'].font = Font(bold=True, size=14)
    
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "Health Score = (Resources at OK Status / Total Resources) × 100%"
    ws[f'A{row}'].font = Font(italic=True, size=9)
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18


# ============================================================================
# SECTION 10: SHEET 8 - COVERAGE ANALYSIS
# ============================================================================

def create_coverage_analysis(wb, styles):
    """Create Coverage Analysis sheet."""
    ws = wb["Coverage_Analysis"]
    
    # Header
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "MONITORING COVERAGE ANALYSIS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:E2')
    cell = ws['A2']
    cell.value = "Identify resources not yet monitored for capacity"
    apply_style(cell, styles['subheader'])
    
    # Coverage metrics
    row = 4
    ws[f'A{row}'] = "Coverage Metric"
    ws[f'B{row}'] = "Target"
    ws[f'C{row}'] = "Actual"
    ws[f'D{row}'] = "Gap"
    ws[f'E{row}'] = "Status"
    
    for col in range(1, 6):
        apply_style(ws.cell(row=row, column=col), styles['column_header'])
    
    # Coverage targets
    row = 5
    metrics = [
        ("Production Systems Coverage", "100%", "", "", ""),
        ("Non-Production Systems Coverage", "90%", "", "", ""),
        ("Overall Coverage", "95%", "", "", ""),
    ]
    
    for metric, target, actual, gap, status in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = target
        apply_style(ws[f'C{row}'], styles['input_cell'])
        apply_style(ws[f'D{row}'], styles['input_cell'])
        apply_style(ws[f'E{row}'], styles['input_cell'])
        row += 1
    
    # Gap details
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "MONITORING GAPS - Resources Not Yet Monitored"
    apply_style(cell, styles['subheader'])
    
    row += 1
    ws[f'A{row}'] = "Resource Name"
    ws[f'B{row}'] = "Resource Type"
    ws[f'C{row}'] = "Reason Not Monitored"
    ws[f'D{row}'] = "Planned Monitoring Date"
    ws[f'E{row}'] = "Responsible Party"
    
    for col in range(1, 6):
        apply_style(ws.cell(row=row, column=col), styles['column_header'])
    
    # Example gap entry
    row += 1
    ws[f'A{row}'] = "[Resource name]"
    ws[f'B{row}'] = "[Type]"
    ws[f'C{row}'] = "[Legacy system, Technical constraint, etc.]"
    ws[f'D{row}'] = "[Date]"
    ws[f'E{row}'] = "[Team/Person]"
    
    for col in range(1, 6):
        apply_style(ws.cell(row=row, column=col), styles['input_cell'])
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25


# ============================================================================
# SECTION 11: SHEET 9 - EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(wb, styles):
    """Create Evidence Register sheet."""
    ws = wb["Evidence_Register"]
    
    # Header
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "EVIDENCE REGISTER"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "Document supporting evidence for capacity utilization assessment"
    apply_style(cell, styles['subheader'])
    
    # Column headers
    row = 4
    headers = [
        "Evidence ID", "Evidence Type", "Description",
        "Location/File Path", "Date Created", "Related Assessment Sheet"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Example evidence entries
    examples = [
        ("A86-MON-001", "Monitoring Dashboard", "Prometheus capacity dashboard screenshot",
         "/evidence/prometheus_dashboard_2026-01.png", "2026-01-11", "All"),
        ("A86-MON-002", "Monitoring Data Export", "CloudWatch metrics export (CSV)",
         "/evidence/cloudwatch_export_2026-01.csv", "2026-01-11", "Compute_Resources"),
        ("A86-CFG-001", "Threshold Configuration", "Alert configuration backup",
         "/evidence/alert_config_2026-01.json", "2026-01-11", "All"),
    ]
    
    row = 5
    for evidence_id, ev_type, desc, location, date, sheet in examples:
        ws[f'A{row}'] = evidence_id
        ws[f'B{row}'] = ev_type
        ws[f'C{row}'] = desc
        ws[f'D{row}'] = location
        ws[f'E{row}'] = date
        ws[f'F{row}'] = sheet
        row += 1
    
    # Apply input styling to remaining rows
    for r in range(row, row + 20):
        for c in range(1, 7):
            apply_style(ws.cell(row=r, column=c), styles['input_cell'])
    
    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25


# ============================================================================
# SECTION 12: SHEET 10 - APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(wb, styles):
    """Create Approval Sign-Off sheet."""
    ws = wb["Approval_Sign_Of"]
    
    # Header
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "ASSESSMENT APPROVAL AND SIGN-OFF"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:D2')
    cell = ws['A2']
    cell.value = "Formal approval of capacity utilization assessment"
    apply_style(cell, styles['subheader'])
    
    # Assessment summary
    row = 4
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "ASSESSMENT SUMMARY"
    apply_style(cell, styles['subheader'])
    
    row += 1
    summary = [
        ("Assessment Date:", "[USER INPUT]"),
        ("Assessment Period:", "[e.g., January 2026]"),
        ("Total Resources Assessed:", "=Threshold_Summary!B10"),
        ("Resources at OK Status:", "=Threshold_Summary!C10"),
        ("Resources at Warning:", "=Threshold_Summary!D10"),
        ("Resources at Critical:", "=Threshold_Summary!E10"),
        ("Capacity Health Score:", "=Threshold_Summary!C12&\"%\""),
    ]
    
    for label, value in summary:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        if value.startswith("="):
            ws[f'B{row}'] = value
        else:
            ws[f'B{row}'] = value
            if "USER INPUT" in value:
                apply_style(ws[f'B{row}'], styles['input_cell'])
        row += 1
    
    # Approval signatures
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "APPROVAL SIGNATURES"
    apply_style(cell, styles['subheader'])
    
    row += 1
    ws[f'A{row}'] = "Role"
    ws[f'B{row}'] = "Name"
    ws[f'C{row}'] = "Signature/Approval"
    ws[f'D{row}'] = "Date"
    
    for col in range(1, 5):
        apply_style(ws.cell(row=row, column=col), styles['column_header'])
    
    # Signature rows
    roles = [
        "Completed By (Capacity Planning Team)",
        "Reviewed By (Infrastructure Manager)",
        "Approved By (IT Operations Manager)",
        "Approved By (CIO/IT Director)",
    ]
    
    row += 1
    for role in roles:
        ws[f'A{row}'] = role
        apply_style(ws[f'B{row}'], styles['input_cell'])
        apply_style(ws[f'C{row}'], styles['input_cell'])
        apply_style(ws[f'D{row}'], styles['input_cell'])
        row += 1
    
    # Notes
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "ASSESSMENT NOTES"
    apply_style(cell, styles['subheader'])
    
    row += 1
    ws.merge_cells(f'A{row}:D{row+5}')
    cell = ws[f'A{row}']
    cell.value = "[Document any significant findings, capacity concerns, or recommendations]"
    apply_style(cell, styles['input_cell'])
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 100
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    print("=" * 80)
    print("ISMS-IMP-A.8.6.1 - Capacity Utilization Assessment Generator")
    print("ISO/IEC 27001:2022 Control A.8.6: Capacity Management")
    print("=" * 80)
    print()
    
    # Create workbook
    print("Creating workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    
    # Create sheets
    print("Generating Instructions & Legend...")
    create_instructions(wb, styles)
    
    print("Generating Compute Resources sheet...")
    create_compute_resources(wb, styles)
    
    print("Generating Storage Resources sheet...")
    create_storage_resources(wb, styles)
    
    print("Generating Network Resources sheet...")
    create_network_resources(wb, styles)
    
    print("Generating Application Resources sheet...")
    create_application_resources(wb, styles)
    
    print("Generating Cloud Resources sheet...")
    create_cloud_resources(wb, styles)
    
    print("Generating Threshold Summary dashboard...")
    create_threshold_summary(wb, styles)
    
    print("Generating Coverage Analysis...")
    create_coverage_analysis(wb, styles)
    
    print("Generating Evidence Register...")
    create_evidence_register(wb, styles)
    
    print("Generating Approval Sign-Off...")
    create_approval_signoff(wb, styles)
    
    # Save workbook
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.8.6.1_Capacity_Utilization_Assessment_{timestamp}.xlsx"
    
    print()
    print(f"Saving workbook: {filename}")
    wb.save(filename)
    
    print()
    print("=" * 80)
    print(f"{CHECK} SUCCESS - Capacity Utilization Assessment Workbook Created")
    print("=" * 80)
    print()
    print(f"Output file: {filename}")
    print()
    print("NEXT STEPS:")
    print("1. Open the workbook in Excel/LibreOffice")
    print("2. Complete yellow-highlighted input cells in each resource sheet")
    print("3. Review auto-calculated utilization percentages and threshold status")
    print("4. Review Threshold_Summary for overall capacity health")
    print("5. Document monitoring gaps in Coverage_Analysis")
    print("6. Add supporting evidence in Evidence_Register")
    print("7. Obtain approvals in Approval_Sign_Off sheet")
    print()
    print("For dashboard integration:")
    print("  • Run normalize_assessment_files_a86.py to prepare for dashboard")
    print("  • Generate dashboard: python3 generate_dashboard_capacity_management.py")
    print()
    print("=" * 80)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
