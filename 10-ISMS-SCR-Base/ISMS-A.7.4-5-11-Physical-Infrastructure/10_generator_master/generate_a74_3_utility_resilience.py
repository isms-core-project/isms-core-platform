#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.7.4.3 - Physical Utility Resilience Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.4: Physical Security Monitoring
Assessment Domain 3 of 4: Utility Infrastructure and Resilience Systems

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific physical security infrastructure, monitoring
systems, and assessment requirements.

Key customization areas:
1. Access control systems and reader types (match your actual hardware)
2. CCTV camera models and recording capabilities (adapt to your installations)
3. Intrusion detection sensor types (specific to your alarm systems)
4. Compliance thresholds (aligned with your security requirements)
5. Incident classification criteria (based on your risk profile)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.7.4 Physical Security Monitoring Framework

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (Python standard library)

Usage:
    python3 generate_a74_1_utility_resilience.py

Output:
    ISMS-IMP-A.7.4.3_Utility_Resilience_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
HOURGLASS = '\u23F3'  # ⏳ Hourglass
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# STYLE DEFINITIONS (A.8.24 PATTERN)
# ============================================================================

def setup_styles():
    """Define all cell styles matching A.8.24 pattern."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# ============================================================================
# INSTRUCTIONS SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.7.4.3 - Physical Utility Resilience Assessment\n"
        "ISO/IEC 27001:2022 - Control A.7.4: Physical Security Monitoring"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.7.4.3"),
        ("Assessment Area", "Physical Utility Resilience Controls"),
        ("Related Policy", "ISMS-POL-A.7.4"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organization", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "1. Complete each worksheet tab for applicable physical security systems.",
        "2. Use dropdown menus for standardized status entries.",
        "3. Fill in yellow-highlighted cells with your facility information.",
        "4. Document all access control readers, CCTV cameras, and intrusion sensors.",
        "5. Track security incidents and response times.",
        "6. Provide evidence location/path for audit traceability.",
        "7. Summary Dashboard auto-calculates compliance metrics.",
        "8. Obtain final approval in the Approval Sign-Off sheet.",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws[f"A{row+1}"] = "STATUS LEGEND"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    legend = [
        ("Symbol", "Status", "Description"),
        (f"{CHECK}", "Compliant", "Fully operational and meets requirements"),
        (f"{WARNING}", "Partial", "Functional but requires attention"),
        (f"{XMARK}", "Non-Compliant", "Not operational or does not meet requirements"),
        ("N/A", "Not Applicable", "Not required for this facility"),
    ]

    row += 2
    for c, h in enumerate(legend[0], start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for sym, status, desc in legend[1:]:
        ws.cell(row=row, column=1, value=sym).border = styles["border"]
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if "Compliant" in status and "Non" not in status:
            s.fill = styles["status_compliant"]["fill"]
        elif "Partial" in status:
            s.fill = styles["status_partial"]["fill"]
        elif "Non-Compliant" in status:
            s.fill = styles["status_noncompliant"]["fill"]
        row += 1

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 50

# ============================================================================
# ACCESS CONTROL SHEET
# ============================================================================

def create_power_infrastructure_sheet(ws, styles):
    """Create Access Control assessment sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Power Infrastructure Assessment\nPolicy Requirement: All facility entry points must have electronic access control"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A3:H3")
    ws["A3"] = "Are UPS systems properly sized with adequate battery runtime and capacity?"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    ws["A4"] = "Response:"
    ws["A4"].font = Font(name="Calibri", bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    dv_response = DataValidation(type="list", formula1='"Yes,No,Not Applicable"', allow_blank=False)
    ws.add_data_validation(dv_response)
    dv_response.add(ws["B4"])

    # Column headers
    columns = {
        "UPS ID": 15,
        "Capacity (kVA)": 18,
        "Load (%)": 12,
        "Battery (%)": 12,
        "Runtime (min)": 15,
        "Status": 15,
        "Evidence Location": 30,
    }

    row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data rows (100 blank rows for data entry)
    for r in range(7, 107):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # Status dropdown (column E)
    dv_status = DataValidation(type="list", 
                               formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
                               allow_blank=False)
    ws.add_data_validation(dv_status)
    for r in range(7, 107):
        dv_status.add(ws.cell(row=r, column=5))

    ws.freeze_panes = "A7"

# ============================================================================
# CCTV SHEET
# ============================================================================

def create_hvac_sheet(ws, styles):
    """Create CCTV Coverage assessment sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "HVAC Systems Assessment\nPolicy Requirement: Video surveillance of all entry/exit points with 90-day retention"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A3:H3")
    ws["A3"] = "Are HVAC systems redundant and properly maintained to meet cooling requirements?"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    ws["A4"] = "Response:"
    ws["A4"].font = Font(name="Calibri", bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    dv_response = DataValidation(type="list", formula1='"Yes,No,Not Applicable"', allow_blank=False)
    ws.add_data_validation(dv_response)
    dv_response.add(ws["B4"])

    columns = {
        "HVAC ID": 15,
        "Capacity (Tons)": 18,
        "Redundancy": 15,
        "Last Service": 15,
        "Status": 15,
        "Evidence Location": 30,
    }

    row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(7, 107):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    dv_status = DataValidation(type="list", 
                               formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
                               allow_blank=False)
    ws.add_data_validation(dv_status)
    for r in range(7, 107):
        dv_status.add(ws.cell(row=r, column=6))

    ws.freeze_panes = "A7"

# ============================================================================
# INTRUSION DETECTION SHEET
# ============================================================================

def create_telecommunications_sheet(ws, styles):
    """Create Intrusion Detection assessment sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Telecommunications Assessment\nPolicy Requirement: Perimeter and critical area monitoring with 24/7 monitoring"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A3:H3")
    ws["A3"] = "Are telecommunications services redundant and meeting SLA targets?"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    ws["A4"] = "Response:"
    ws["A4"].font = Font(name="Calibri", bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    dv_response = DataValidation(type="list", formula1='"Yes,No,Not Applicable"', allow_blank=False)
    ws.add_data_validation(dv_response)
    dv_response.add(ws["B4"])

    columns = {
        "ISP ID": 15,
        "Provider": 20,
        "Bandwidth (Mbps)": 18,
        "SLA (%)": 12,
        "Actual (%)": 12,
        "Status": 15,
        "Evidence Location": 30,
    }

    row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(7, 107):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    dv_status = DataValidation(type="list", 
                               formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
                               allow_blank=False)
    ws.add_data_validation(dv_status)
    for r in range(7, 107):
        dv_status.add(ws.cell(row=r, column=6))

    ws.freeze_panes = "A7"

# Incidents sheet not applicable for utility resilience

def create_incidents_sheet_REMOVED(ws, styles):
    """Create Security Incidents tracking sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Physical Security Incidents\nPolicy Requirement: All incidents logged and investigated with <15min response time"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A3:H3")
    ws["A3"] = "Log all physical security incidents including unauthorized access attempts, tailgating, and alarm triggers."
    ws["A3"].font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    columns = {
        "Incident ID": 15,
        "Date": 12,
        "Type": 20,
        "Severity": 12,
        "Location": 20,
        "Response Time (min)": 18,
        "Resolved": 12,
        "Notes": 35,
    }

    row = 5
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(6, 106):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    dv_type = DataValidation(type="list", 
                             formula1='"Tailgating,Unauthorized Access,Alarm Trigger,Lost Badge,Forced Entry,Other"',
                             allow_blank=False)
    ws.add_data_validation(dv_type)
    for r in range(6, 106):
        dv_type.add(ws.cell(row=r, column=3))

    dv_severity = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=False)
    ws.add_data_validation(dv_severity)
    for r in range(6, 106):
        dv_severity.add(ws.cell(row=r, column=4))

    ws.freeze_panes = "A6"

# ============================================================================
# SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with metrics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "ACCESS MONITORING SUMMARY DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Period:"
    ws["A3"].font = Font(name="Calibri", bold=True)
    ws["B3"].fill = styles["input_cell"]["fill"]
    ws["B3"].border = styles["border"]

    ws["A5"] = "Compliance Metrics"
    ws["A5"].font = Font(name="Calibri", size=12, bold=True)

    headers = ["Metric", "Count", f"{CHECK} Compliant", f"{WARNING} Partial", f"{XMARK} Non-Compliant", "N/A", "% Compliant"]
    row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    areas = [
        ("Power Infrastructure", "Power Infrastructure", 6),
        ("HVAC Systems", "HVAC", 5),
        ("Telecommunications", "Telecommunications", 6),
    ]

    row += 1
    start_data_row = row
    for label, sheet, status_col in areas:
        ws.cell(row=row, column=1, value=label)
        
        status_col_letter = get_column_letter(status_col)
        status_range = f"'{sheet}'!{status_col_letter}7:{status_col_letter}106"

        ws.cell(row=row, column=2, value=f'=COUNTA({status_range})')
        ws.cell(row=row, column=3, value=f'=COUNTIF({status_range},"{CHECK}*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({status_range},"{WARNING}*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({status_range},"{XMARK}*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({status_range},"N/A")')
        ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{row-1})")
        cell.font = Font(name="Calibri", bold=True)

    total_pct = ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
    total_pct.font = Font(name="Calibri", bold=True, color="0000FF", size=12)

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 18

    # Incident metrics
    row += 3
    ws[f"A{row}"] = "Incident Response Metrics"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)
    
    row += 1
    ws[f"A{row}"] = "Total Incidents (Last Period):"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws[f"B{row}"] = "=COUNTA(Incidents!A6:A105)"
    
    row += 1
    ws[f"A{row}"] = "Average Response Time (minutes):"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws[f"B{row}"] = "=AVERAGE(Incidents!F6:F105)"

# ============================================================================
# EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID",
        "Assessment Area",
        "Evidence Type",
        "Description",
        "Location/Path",
        "Date Collected",
        "Collected By",
        "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Access log,Video footage,Test report,Maintenance record,Audit log,Other"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_type)

    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_ver)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EV-{r-4:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"

# ============================================================================
# APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create Approval and Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Summary"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    summary_fields = [
        ("Assessment Document", "ISMS-IMP-A.7.4.3 - Utility Resilience Assessment"),
        ("Assessment Period", ""),
        ("Overall Compliance Rate", "='Summary Dashboard'!G9"),
        ("Assessment Status", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Approval Workflow"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    approval_headers = ["Role", "Name", "Signature", "Date"]
    row += 3
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    roles = ["Assessor", "Security Manager", "ISO Reviewer", "CISO Approver"]
    row += 1
    for role in roles:
        ws.cell(row=row, column=1, value=role).font = Font(name="Calibri", bold=True)
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15

# ============================================================================
# MAIN WORKBOOK GENERATION
# ============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)
    
    styles = setup_styles()
    
    # Create all sheets
    ws = wb.create_sheet("Instructions & Legend", 0)
    create_instructions_sheet(ws, styles)
    
    ws = wb.create_sheet("Power Infrastructure")
    create_power_infrastructure_sheet(ws, styles)
    
    ws = wb.create_sheet("HVAC")
    create_hvac_sheet(ws, styles)
    
    ws = wb.create_sheet("Telecommunications")
    create_telecommunications_sheet(ws, styles)
    
    # Incidents sheet not needed
    
    ws = wb.create_sheet("Summary Dashboard")
    create_summary_dashboard(ws, styles)
    
    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)
    
    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_signoff(ws, styles)
    
    return wb

# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == "__main__":
    print(f"\n{'='*70}")
    print("ISMS Physical Utility Resilience Assessment (A.7.4.3)")
    print(f"{'='*70}\n")
    
    try:
        wb = create_workbook()
        filename = f"ISMS-IMP-A.7.4.3_Utility_Resilience_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)
        
        print(f"{CHECK} SUCCESS: {filename}")
        print(f"  {BULLET} 7 professional worksheets created")
        print(f"  {BULLET} A.8.24 styling applied (navy headers, yellow inputs)")
        print(f"  {BULLET} 100 data rows per assessment sheet")
        print(f"  {BULLET} Automated compliance dashboard with formulas")
        print(f"  {BULLET} Data validations and freeze panes configured")
        print(f"  {BULLET} Evidence register with audit traceability")
        print(f"  {BULLET} 4-level approval workflow")
        print(f"\n{'='*70}\n")
        
    except Exception as e:
        print(f"\n{XMARK} ERROR: Failed to generate workbook")
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
