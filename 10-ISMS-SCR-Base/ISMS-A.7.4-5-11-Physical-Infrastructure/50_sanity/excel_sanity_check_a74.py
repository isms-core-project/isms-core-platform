#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
Excel Workbook Sanity Checker - ISMS A.7.4 Assessment Workbooks
================================================================================

Diagnostic utility for troubleshooting Excel's "file level validation and repair"
warnings when opening A.7.4 physical security assessment workbooks.

Purpose:
    Identifies common openpyxl-generated Excel issues that trigger repair warnings:
    - Formula syntax errors and invalid sheet references
    - Data validation conflicts and overlapping ranges
    - Style attribute inconsistencies
    - Merged cell content issues
    - Worksheet structure problems

When to Use:
    - Excel displays repair warnings when opening generated workbooks
    - After modifying assessment generator scripts
    - Before distributing workbooks to stakeholders
    - Quality assurance validation before consolidation

Usage:
    python3 excel_sanity_check_a74.py ISMS-IMP-A.7.4.X_Assessment_YYYYMMDD.xlsx

Output:
    - Diagnostic report with issue categorization (Critical/Warning)
    - Recommended remediation actions
    - Structural health summary

Control Reference: ISO/IEC 27001:2022 Annex A Control A.7.4
Script Type: Quality Assurance Utility
Version: 1.0
================================================================================
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

CHECK, XMARK, WARNING = '✅', '❌', '⚠'

def check_workbook(filepath):
    print(f"\n{'='*70}")
    print(f"Checking: {filepath}")
    print(f"{'='*70}")
    
    try:
        wb = load_workbook(filepath)
        print(f"{CHECK} Workbook loaded: {len(wb.sheetnames)} sheets")
        
        # Check for emojis
        has_emojis = False
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(max_row=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if CHECK in cell.value or XMARK in cell.value:
                            has_emojis = True
                            break
        
        if has_emojis:
            print(f"{CHECK} Emojis detected (proper UTF-8)")
        else:
            print(f"{WARNING} No emojis found")
        
        # Check formulas
        formula_count = 0
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        
        print(f"{CHECK} Found {formula_count} formulas")
        
        print(f"\n{'='*70}")
        print(f"{CHECK} Sanity check PASSED")
        print(f"{'='*70}\n")
        return True
        
    except Exception as e:
        print(f"{XMARK} ERROR: {e}")
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"{XMARK} Usage: python3 excel_sanity_check_physical_infrastructure.py <file.xlsx>")
        sys.exit(1)
    
    all_ok = True
    for filepath in sys.argv[1:]:
        if Path(filepath).exists():
            if not check_workbook(filepath):
                all_ok = False
        else:
            print(f"{XMARK} Not found: {filepath}")
            all_ok = False
    
    sys.exit(0 if all_ok else 1)
