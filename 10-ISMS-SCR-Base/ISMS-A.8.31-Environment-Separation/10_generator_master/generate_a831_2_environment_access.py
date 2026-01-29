#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
Environment Access Control Assessment Generator - ISMS A.8.31 Domain 2
================================================================================

Generates Excel assessment workbook for evaluating environment access control
compliance per ISO/IEC 27001:2022 Control A.8.31.

**Purpose:**
Creates structured assessment for environment access restrictions including:
- Developer production access verification (should be zero/break-glass only)
- Environment-specific access matrices (who can access which environment)
- Production credential management (PAM vault usage)
- Cross-environment access attempt monitoring

**Assessment Focus:**
- Production access restrictions (developers should NOT have production access)
- Environment access provisioning (role-based per environment)
- Emergency access procedures (break-glass workflows, time-limited)
- Access monitoring and alerting (unauthorized cross-environment attempts)

**Usage:**
    python3 generate_a831_2_environment_access.py
    
    Output: ISMS_IMP_A_8_31_2_Environment_Access_Control_YYYYMMDD.xlsx

**Workbook Structure:**
- Executive Summary (access compliance status, policy violations)
- Gap Analysis (unauthorized access, missing restrictions)
- User Environment Matrix (user → environment access mapping)
- Production Access Audit (developers with prod access - should be 0)
- Access Provisioning (how access is granted per environment)
- Emergency Access Log (break-glass usage, post-incident reviews)
- Credential Management (env-specific credentials, PAM integration)
- Evidence Register (access logs, IAM policies, RBAC configs)
- Action Items (access revocations, policy enforcement)

**Integration:**
- Consolidates into A.8.31 Compliance Dashboard
- Links to A.5.15-16-18 (IAM), A.8.2-3-5 (Auth-PAM)
- Supports incident response (A.5.24-27) for unauthorized access

Control Reference: ISO/IEC 27001:2022 Annex A Control A.8.31
Script Type: Assessment Workbook Generator
Version: 1.0
================================================================================
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# WORKBOOK CREATION
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions_Legend",
        "User_Environment_Access_Matrix",
        "Developer_Production_Access",
        "Production_Credential_Audit",
        "Cross_Environment_Access_Log",
        "Break_Glass_Access_Log",
        "MFA_Enforcement",
        "Access_Compliance_Scorecard",
        "Evidence_Register",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "formula_cell": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
        "status_green": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_red": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "critical_violation": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style to cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], "color") else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, "rgb") else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, "rgb") else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# DATA VALIDATIONS
# ============================================================================

def create_base_validations(ws):
    """Create data validation objects."""
    validations = {
        "yes_no": DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        "yes_no_na": DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,➖ N/A"',
            allow_blank=False
        ),
        "access_level": DataValidation(
            type="list",
            formula1='"Full (CRUD),Read/Write,Read-only,No Access,Break-Glass Only"',
            allow_blank=False
        ),
        "compliance_status": DataValidation(
            type="list",
            formula1='"✅ Compliant,❌ Non-Compliant,⚠️ Partial,📋 Not Assessed"',
            allow_blank=False
        ),
        "user_role": DataValidation(
            type="list",
            formula1='"Developer,QA Engineer,DevOps Engineer,Operations Engineer,Security Analyst,Manager,Auditor,Other"',
            allow_blank=False
        ),
        "environment_type": DataValidation(
            type="list",
            formula1='"Development,Testing,Staging,Production"',
            allow_blank=False
        ),
        "severity": DataValidation(
            type="list",
            formula1='"🔴 Critical,🟠 High,🟡 Medium,🟢 Low"',
            allow_blank=False
        ),
    }
    
    for val in validations.values():
        ws.add_data_validation(val)
    
    return validations


# ============================================================================
# INSTRUCTIONS SHEET
# ============================================================================

def create_instructions_sheet(wb, styles):
    """Create Instructions & Legend sheet."""
    ws = wb["Instructions_Legend"]
    
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "ISMS-IMP-A.8.31.2 – Environment Access Control Assessment"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells("A2:F2")
    cell = ws["A2"]
    cell.value = "ISO/IEC 27001:2022 - Control A.8.31: Environment Access Control"
    apply_style(cell, styles["subheader"])
    ws.row_dimensions[2].height = 30
    
    row = 4
    info_items = [
        ("Document ID:", "ISMS-IMP-A.8.31.2"),
        ("Assessment Area:", "Environment Access Control & Restrictions"),
        ("Related Policy:", "ISMS-POL-A.8.31-S2.2"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Quarterly"),
    ]
    
    for label, value in info_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "USER INPUT" in value:
            apply_style(cell, styles["input_cell"])
        row += 1
    
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws.cell(row=row, column=1, value="How to Use This Workbook")
    apply_style(cell, styles["subheader"])
    
    row += 1
    instructions = [
        "1. Complete User_Environment_Access_Matrix - document who has access to which environment",
        "2. ⚠️ CRITICAL: Complete Developer_Production_Access - verify ZERO developers have production access",
        "3. Audit Production_Credential_Audit - verify all production credentials in PAM vault",
        "4. Review Cross_Environment_Access_Log - check for unauthorized access attempts",
        "5. Document Break_Glass_Access_Log - all emergency production access instances",
        "6. Verify MFA_Enforcement - confirm MFA required for production",
        "7. Complete Access_Compliance_Scorecard - calculate overall compliance",
        "8. Maintain Evidence_Register for audit traceability",
        "9. Obtain final approval and sign-off",
    ]
    
    for instruction in instructions:
        ws.cell(row=row, column=1, value=instruction)
        row += 1
    
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws.cell(row=row, column=1, value="CRITICAL REQUIREMENT")
    apply_style(cell, styles["critical_violation"])
    
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="🚨 ZERO developers shall have production access (except via documented break-glass)")
    
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws.cell(row=row, column=1, value="Status Legend")
    apply_style(cell, styles["subheader"])
    
    row += 1
    legend = [
        ("✅", "Compliant / Authorized", "Green"),
        ("❌", "Non-Compliant / Unauthorized", "Red"),
        ("⚠️", "Partial / Warning", "Yellow"),
        ("🔴", "Critical Violation", "Dark Red"),
        ("➖", "N/A / Not Applicable", "Gray"),
    ]
    
    ws.cell(row=row, column=1, value="Symbol").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Status").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Color").font = Font(bold=True)
    row += 1
    
    for symbol, status, color in legend:
        ws.cell(row=row, column=1, value=symbol)
        ws.cell(row=row, column=2, value=status)
        ws.cell(row=row, column=3, value=color)
        row += 1
    
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20


# ============================================================================
# USER ENVIRONMENT ACCESS MATRIX
# ============================================================================

def create_user_access_matrix_sheet(wb, styles):
    """Create User-Environment Access Matrix."""
    ws = wb["User_Environment_Access_Matrix"]
    
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "USER → ENVIRONMENT ACCESS MATRIX"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = "Document access permissions for each user per environment"
    apply_style(cell, styles["subheader"])
    
    row = 4
    headers = [
        ("A", "User ID / Email", 30),
        ("B", "Role", 25),
        ("C", "Development Access", 25),
        ("D", "Testing Access", 25),
        ("E", "Staging Access", 25),
        ("F", "Production Access", 25),
        ("G", "Compliance Status", 20),
        ("H", "Notes", 40),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    validations = create_base_validations(ws)
    validations["user_role"].add("B5:B100")
    validations["access_level"].add("C5:F100")
    validations["compliance_status"].add("G5:G100")
    
    row = 5
    sample_data = [
        ("dev1@example.com", "Developer", "Full (CRUD)", "Read/Write", "Read-only", "❌ No Access", "✅ Compliant", "Proper access separation"),
        ("dev2@example.com", "Developer", "Full (CRUD)", "Read/Write", "Read-only", "❌ No Access", "✅ Compliant", "No prod access (expected)"),
        ("qa1@example.com", "QA Engineer", "Read-only", "Full (CRUD)", "Read/Write", "❌ No Access", "✅ Compliant", "Testing environment focus"),
        ("ops1@example.com", "Operations Engineer", "Read-only", "Read-only", "Full (CRUD)", "Full (CRUD)", "✅ Compliant", "Prod access via PAM vault"),
        ("ops2@example.com", "Operations Engineer", "❌ No Access", "❌ No Access", "Full (CRUD)", "Full (CRUD)", "✅ Compliant", "Production-focused role"),
    ]
    
    for data in sample_data:
        for idx, value in enumerate(data):
            cell = ws.cell(row=row, column=idx+1, value=value)
            apply_style(cell, styles["input_cell"])
        row += 1


# ============================================================================
# DEVELOPER PRODUCTION ACCESS (CRITICAL CHECK)
# ============================================================================

def create_developer_prod_access_sheet(wb, styles):
    """Create Developer Production Access critical check sheet."""
    ws = wb["Developer_Production_Access"]
    
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "🚨 DEVELOPER PRODUCTION ACCESS CHECK (CRITICAL)"
    apply_style(cell, styles["critical_violation"])
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells("A2:G2")
    cell = ws["A2"]
    cell.value = "TARGET: ZERO developers with production access | Any violations = CRITICAL FINDING"
    apply_style(cell, styles["subheader"])
    ws.row_dimensions[2].height = 30
    
    row = 4
    headers = [
        ("A", "Developer ID", 30),
        ("B", "Production Account/Sub", 30),
        ("C", "Production Access?", 20),
        ("D", "Access Type", 25),
        ("E", "Justification", 40),
        ("F", "Violation Severity", 20),
        ("G", "Remediation Action", 40),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    validations = create_base_validations(ws)
    validations["yes_no"].add("C5:C100")
    validations["severity"].add("F5:F100")
    
    row = 5
    ws.cell(row=row, column=1, value="dev1@example.com")
    ws.cell(row=row, column=2, value="AWS Prod Account: 444444444444")
    ws.cell(row=row, column=3, value="No")
    ws.cell(row=row, column=4, value="N/A")
    ws.cell(row=row, column=5, value="No production access (expected)")
    ws.cell(row=row, column=6, value="🟢 Low")
    ws.cell(row=row, column=7, value="N/A - Compliant")
    
    for col in range(1, 8):
        apply_style(ws.cell(row=row, column=col), styles["input_cell"])
    
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="✅ If ALL developers show 'No' for Production Access → COMPLIANT")
    apply_style(cell, styles["status_green"])
    
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="🔴 If ANY developer shows 'Yes' for Production Access → CRITICAL VIOLATION")
    apply_style(cell, styles["critical_violation"])


# ============================================================================
# PRODUCTION CREDENTIAL AUDIT
# ============================================================================

def create_production_credential_audit_sheet(wb, styles):
    """Create Production Credential Audit sheet."""
    ws = wb["Production_Credential_Audit"]
    
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "PRODUCTION CREDENTIAL AUDIT"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = "Verify all production credentials stored in PAM vault and rotated regularly"
    apply_style(cell, styles["subheader"])
    
    row = 4
    headers = [
        ("A", "Credential Type", 30),
        ("B", "System/Service", 30),
        ("C", "Stored in PAM Vault?", 20),
        ("D", "Vault Location", 35),
        ("E", "Last Rotated", 20),
        ("F", "Rotation Schedule", 20),
        ("G", "Compliance", 20),
        ("H", "Notes", 40),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    validations = create_base_validations(ws)
    validations["yes_no"].add("C5:C100")
    validations["compliance_status"].add("G5:G100")
    
    row = 5
    sample_data = [
        ("Database Admin Password", "prod-db.example.com", "Yes", "CyberArk/Prod/DB/AdminPwd", "2026-01-01", "90 days", "✅ Compliant", "Auto-rotation enabled"),
        ("AWS Root Account", "AWS Prod: 444444444444", "Yes", "CyberArk/Prod/AWS/Root", "2025-12-15", "Never (not used)", "✅ Compliant", "MFA enforced, never used"),
        ("SSH Private Key", "prod-app-servers", "Yes", "CyberArk/Prod/SSH/Keys", "2025-11-20", "180 days", "✅ Compliant", "Key-based auth"),
        ("API Key (Stripe)", "Payment gateway", "Yes", "AWS Secrets Manager", "2026-01-05", "90 days", "✅ Compliant", "Live API key"),
    ]
    
    for data in sample_data:
        for idx, value in enumerate(data):
            cell = ws.cell(row=row, column=idx+1, value=value)
            apply_style(cell, styles["input_cell"])
        row += 1


# ============================================================================
# CROSS ENVIRONMENT ACCESS LOG
# ============================================================================

def create_cross_environment_access_sheet(wb, styles):
    """Create Cross-Environment Access Log sheet."""
    ws = wb["Cross_Environment_Access_Log"]
    
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "CROSS-ENVIRONMENT ACCESS ATTEMPTS LOG"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = "Log unauthorized cross-environment access attempts (should be blocked)"
    apply_style(cell, styles["subheader"])
    
    row = 4
    headers = [
        ("A", "Date/Time", 20),
        ("B", "User ID", 30),
        ("C", "Source Environment", 20),
        ("D", "Target Environment", 20),
        ("E", "Attempted Action", 30),
        ("F", "Result", 20),
        ("G", "Alert Generated?", 20),
        ("H", "Investigation Notes", 40),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    validations = create_base_validations(ws)
    validations["environment_type"].add("C5:D100")
    validations["yes_no"].add("G5:G100")
    
    row = 5
    sample_data = [
        ("2026-01-10 14:23:00", "dev1@example.com", "Development", "Production", "SSH to prod-app01", "❌ BLOCKED", "Yes", "Firewall rule blocked access (expected)"),
        ("2026-01-09 09:15:00", "dev2@example.com", "Development", "Production", "AWS CLI describe-instances", "❌ DENIED", "Yes", "IAM policy denied (expected)"),
    ]
    
    for data in sample_data:
        for idx, value in enumerate(data):
            cell = ws.cell(row=row, column=idx+1, value=value)
            apply_style(cell, styles["input_cell"])
        row += 1


# ============================================================================
# BREAK GLASS ACCESS LOG
# ============================================================================

def create_breakglass_access_sheet(wb, styles):
    """Create Break-Glass Access Log sheet."""
    ws = wb["Break_Glass_Access_Log"]
    
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "BREAK-GLASS EMERGENCY ACCESS LOG"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:J2")
    cell = ws["A2"]
    cell.value = "Document all instances of emergency developer access to production"
    apply_style(cell, styles["subheader"])
    
    row = 4
    headers = [
        ("A", "Incident ID", 15),
        ("B", "Date/Time Activated", 20),
        ("C", "Developer ID", 30),
        ("D", "Approved By", 25),
        ("E", "Reason/Justification", 40),
        ("F", "Duration (hours)", 15),
        ("G", "Date/Time Revoked", 20),
        ("H", "Post-Incident Review", 30),
        ("I", "Compliance", 20),
        ("J", "Notes", 40),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    validations = create_base_validations(ws)
    validations["compliance_status"].add("I5:I100")
    
    row = 5
    sample_data = [
        ("INC-2026-001", "2026-01-05 02:30", "senior-dev@example.com", "ops-manager@example.com", "Critical production outage - database corruption", "4", "2026-01-05 06:30", "Completed 2026-01-06", "✅ Compliant", "Proper procedure followed"),
    ]
    
    for data in sample_data:
        for idx, value in enumerate(data):
            cell = ws.cell(row=row, column=idx+1, value=value)
            apply_style(cell, styles["input_cell"])
        row += 1


# ============================================================================
# MFA ENFORCEMENT
# ============================================================================

def create_mfa_enforcement_sheet(wb, styles):
    """Create MFA Enforcement sheet."""
    ws = wb["MFA_Enforcement"]
    
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "MFA ENFORCEMENT VERIFICATION"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:G2")
    cell = ws["A2"]
    cell.value = "Verify MFA required for production access (all users)"
    apply_style(cell, styles["subheader"])
    
    row = 4
    headers = [
        ("A", "User ID", 30),
        ("B", "Production Access?", 20),
        ("C", "MFA Enabled?", 20),
        ("D", "MFA Type", 25),
        ("E", "Last MFA Check", 20),
        ("F", "Compliance", 20),
        ("G", "Evidence", 40),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    validations = create_base_validations(ws)
    validations["yes_no"].add("B5:C100")
    validations["compliance_status"].add("F5:F100")
    
    row = 5
    sample_data = [
        ("ops1@example.com", "Yes", "Yes", "Hardware Token (YubiKey)", "2026-01-11", "✅ Compliant", "IAM user MFA device ARN: ...mfa/ops1"),
        ("ops2@example.com", "Yes", "Yes", "Virtual MFA (Google Authenticator)", "2026-01-11", "✅ Compliant", "IAM user MFA device ARN: ...mfa/ops2"),
    ]
    
    for data in sample_data:
        for idx, value in enumerate(data):
            cell = ws.cell(row=row, column=idx+1, value=value)
            apply_style(cell, styles["input_cell"])
        row += 1


# ============================================================================
# COMPLIANCE SCORECARD
# ============================================================================

def create_compliance_scorecard_sheet(wb, styles):
    """Create Access Compliance Scorecard."""
    ws = wb["Access_Compliance_Scorecard"]
    
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = "ACCESS COMPLIANCE SCORECARD"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:E2")
    cell = ws["A2"]
    cell.value = "Summary of access control compliance metrics"
    apply_style(cell, styles["subheader"])
    
    row = 4
    headers = [
        ("A", "Metric", 40),
        ("B", "Current Value", 20),
        ("C", "Target Value", 20),
        ("D", "Status", 20),
        ("E", "Notes", 40),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    row = 5
    metrics = [
        ("🎯 Developers with Production Access", "0", "0", "✅ Target Met", "Critical requirement achieved"),
        ("Production Users with MFA", "100%", "100%", "✅ Target Met", "All production accounts MFA-enabled"),
        ("Break-Glass Usage (Past Month)", "1", "< 5", "✅ Acceptable", "Within acceptable range"),
        ("Cross-Environment Access Attempts", "2", "0 (blocked)", "✅ Acceptable", "All attempts blocked by controls"),
        ("Production Credentials in PAM Vault", "100%", "100%", "✅ Target Met", "All prod credentials vaulted"),
        ("Access Reviews Completed On Time", "100%", "100%", "✅ Target Met", "Monthly reviews up to date"),
        ("Dormant Accounts (90+ days)", "0", "0", "✅ Target Met", "No inactive accounts"),
    ]
    
    for metric, current, target, status, notes in metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=current)
        ws.cell(row=row, column=3, value=target)
        cell = ws.cell(row=row, column=4, value=status)
        ws.cell(row=row, column=5, value=notes)
        
        if "✅" in status:
            apply_style(cell, styles["status_green"])
        elif "❌" in status:
            apply_style(cell, styles["status_red"])
        
        for col in range(1, 6):
            apply_style(ws.cell(row=row, column=col), styles["input_cell"])
        
        row += 1


# ============================================================================
# EVIDENCE REGISTER
# ============================================================================

def create_evidence_register_sheet(wb, styles):
    """Create Evidence Register sheet."""
    ws = wb["Evidence_Register"]
    
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "EVIDENCE REGISTER"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:G2")
    cell = ws["A2"]
    cell.value = "Maintain audit trail of all supporting evidence"
    apply_style(cell, styles["subheader"])
    
    row = 4
    headers = [
        ("A", "Evidence ID", 15),
        ("B", "Evidence Type", 30),
        ("C", "Description", 40),
        ("D", "Related Requirement", 30),
        ("E", "File Location", 40),
        ("F", "Collected Date", 15),
        ("G", "Collected By", 20),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    row = 5
    evidence = [
        ("EV-ACC-001", "IAM Policy Export", "AWS IAM policies showing dev/prod separation", "Production Access Control", "/evidence/aws_iam_policies.json", "2026-01-11", "IAM Admin"),
        ("EV-ACC-002", "User Access Matrix", "Complete user-environment access matrix", "Access Matrix", "/evidence/access_matrix_20260111.xlsx", "2026-01-11", "Security Analyst"),
        ("EV-ACC-003", "MFA Enforcement Policy", "Conditional access policy requiring MFA for prod", "MFA Enforcement", "/evidence/conditional_access_mfa.png", "2026-01-11", "IAM Admin"),
        ("EV-ACC-004", "Break-Glass Log", "Past 12 months break-glass access instances", "Break-Glass Procedure", "/evidence/breakglass_log_2025.xlsx", "2026-01-11", "Security Manager"),
        ("EV-ACC-005", "Production Access Log", "CloudTrail logs showing production access", "Access Monitoring", "/evidence/cloudtrail_prod_access.csv", "2026-01-11", "Security Analyst"),
    ]
    
    for data in evidence:
        for idx, value in enumerate(data):
            cell = ws.cell(row=row, column=idx+1, value=value)
            apply_style(cell, styles["input_cell"])
        row += 1


# ============================================================================
# APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff_sheet(wb, styles):
    """Create Approval Sign-Off sheet."""
    ws = wb["Approval_Sign_Off"]
    
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = "APPROVAL & SIGN-OFF"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:E2")
    cell = ws["A2"]
    cell.value = "Formal approval of access control assessment"
    apply_style(cell, styles["subheader"])
    
    row = 4
    ws.cell(row=row, column=1, value="Assessment Summary:").font = Font(bold=True)
    row += 1
    
    summary_items = [
        ("Total Users Assessed:", "[COUNT]"),
        ("Users with Production Access:", "[COUNT]"),
        ("Developers with Production Access:", "0 (TARGET)"),
        ("Production Users with MFA:", "[COUNT] (100% TARGET)"),
        ("Break-Glass Instances (Past 12 Months):", "[COUNT]"),
        ("Critical Findings:", "[COUNT]"),
    ]
    
    for label, value in summary_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value="Approval Signatures")
    apply_style(cell, styles["subheader"])
    
    row += 1
    headers = ["Role", "Name", "Signature", "Date", "Comments"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=header)
        apply_style(cell, styles["column_header"])
    
    row += 1
    roles = [
        "Assessment Lead",
        "CISO",
        "IT Operations Manager",
        "IAM Administrator",
        "Information Security Manager"
    ]
    
    for role in roles:
        ws.cell(row=row, column=1, value=role)
        for col in range(1, 6):
            apply_style(ws.cell(row=row, column=col), styles["input_cell"])
        row += 1
    
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 40


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Generate the assessment workbook."""
    print("=" * 80)
    print("ISMS-IMP-A.8.31.2 - Environment Access Control Assessment Generator")
    print("ISO/IEC 27001:2022 Control A.8.31")
    print("=" * 80)
    
    print("\nCreating workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    
    print("Generating Instructions & Legend sheet...")
    create_instructions_sheet(wb, styles)
    
    print("Generating User-Environment Access Matrix sheet...")
    create_user_access_matrix_sheet(wb, styles)
    
    print("Generating Developer Production Access sheet (CRITICAL)...")
    create_developer_prod_access_sheet(wb, styles)
    
    print("Generating Production Credential Audit sheet...")
    create_production_credential_audit_sheet(wb, styles)
    
    print("Generating Cross-Environment Access Log sheet...")
    create_cross_environment_access_sheet(wb, styles)
    
    print("Generating Break-Glass Access Log sheet...")
    create_breakglass_access_sheet(wb, styles)
    
    print("Generating MFA Enforcement sheet...")
    create_mfa_enforcement_sheet(wb, styles)
    
    print("Generating Access Compliance Scorecard sheet...")
    create_compliance_scorecard_sheet(wb, styles)
    
    print("Generating Evidence Register sheet...")
    create_evidence_register_sheet(wb, styles)
    
    print("Generating Approval Sign-Off sheet...")
    create_approval_signoff_sheet(wb, styles)
    
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.8.31.2_Environment_Access_Control_Assessment_{timestamp}.xlsx"
    
    print(f"\nSaving workbook: {filename}")
    wb.save(filename)
    
    print("=" * 80)
    print("✅ SUCCESS!")
    print("=" * 80)
    print(f"\nGenerated: {filename}")
    print("\nNext Steps:")
    print("1. Complete User_Environment_Access_Matrix")
    print("2. 🚨 CRITICAL: Verify Developer_Production_Access shows ZERO developers")
    print("3. Audit Production_Credential_Audit (all creds in PAM vault)")
    print("4. Review Break_Glass_Access_Log")
    print("5. Verify MFA_Enforcement (100% for production)")
    print("6. Complete Access_Compliance_Scorecard")
    print("\n" + "=" * 80)
    

if __name__ == "__main__":
    main()
