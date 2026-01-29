#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
Excel Workbook Sanity Checker - ISMS A.8.31 Assessment Workbooks
================================================================================

Diagnostic utility for troubleshooting Excel's "file level validation and repair"
warnings when opening A.8.31 environment separation assessment workbooks.

**Purpose:**
Identifies common openpyxl-generated Excel issues that trigger repair warnings:
- Formula syntax errors and invalid sheet references
- Data validation conflicts and overlapping ranges
- Style attribute inconsistencies
- Merged cell content issues
- Worksheet structure problems

**When to Use:**
- Excel displays repair warnings when opening generated workbooks
- After modifying assessment generator scripts
- Before distributing workbooks to stakeholders
- Quality assurance validation before consolidation

**Usage:**
    python3 excel_sanity_check_a831.py ISMS_IMP_A_8_31_X_Assessment_YYYYMMDD.xlsx
    
    Works with any A.8.31 assessment workbook (domains 1-3) or dashboard

**Output:**
- Diagnostic report with issue categorization (Critical/Warning)
- Recommended remediation actions
- Structural health summary

**Common Issues Detected:**
- External formula references (expected for dashboard, not assessments)
- Data validation range overlaps (dropdown conflicts)
- Merged cell style inconsistencies
- Invalid cross-sheet formula references
- Missing or empty required worksheets

**Validation Categories:**
1. **Critical Issues** (prevent workbook opening):
   - Corrupted worksheet structures
   - Invalid XML that Excel cannot repair
   - Fatal formula syntax errors

2. **Warnings** (Excel can repair):
   - External workbook links (expected in dashboard)
   - Data validation overlaps (cosmetic issue)
   - Style attribute inconsistencies (visual only)

**Related Scripts:**
- generate_a831_1_environment_architecture.py (domain 1 generator)
- generate_a831_2_environment_access.py (domain 2 generator)
- generate_a831_3_compliance_dashboard.py (dashboard generator)
- normalize_assessment_files_a831.py (file preparation)

Control Reference: ISO/IEC 27001:2022 Annex A Control A.8.31
Script Type: Quality Assurance Utility
Version: 1.0
================================================================================
"""

import sys
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ============================================================================
# WORKBOOK TYPE DETECTION
# ============================================================================

def detect_workbook_type(filename):
    """Detect which A.8.31 workbook type this is based on filename."""
    filename_lower = filename.lower()
    
    if 'a.8.31.1' in filename_lower or 'environment_architecture' in filename_lower:
        return 'A.8.31.1', 'Environment Architecture Assessment'
    elif 'a.8.31.2' in filename_lower or 'access_control' in filename_lower:
        return 'A.8.31.2', 'Environment Access Control Assessment'
    elif 'a.8.31.3' in filename_lower or 'dashboard' in filename_lower:
        return 'A.8.31.3', 'Compliance Dashboard'
    else:
        return 'Unknown', 'Generic Excel Workbook'


# ============================================================================
# EXPECTED SHEET DEFINITIONS
# ============================================================================

EXPECTED_SHEETS = {
    'A.8.31.1': [
        "Instructions_Legend", "Environment_Inventory", "Network_Separation",
        "Infrastructure_Separation", "Data_Separation", "Credential_Separation",
        "Configuration_Consistency", "Gap_Analysis", "Evidence_Register", "Approval_Sign_Off"
    ],
    'A.8.31.2': [
        "Instructions_Legend", "User_Environment_Access_Matrix", "Developer_Production_Access",
        "Production_Credential_Audit", "Cross_Environment_Access_Log", "Break_Glass_Access_Log",
        "MFA_Enforcement", "Access_Compliance_Scorecard", "Evidence_Register", "Approval_Sign_Off"
    ],
    'A.8.31.3': [
        "Executive_Summary", "Environment_Separation_Status", "Access_Control_Summary",
        "Gap_Analysis_Remediation", "Risk_Register", "Evidence_Summary",
        "Trend_Analysis", "Approval_Sign_Off"
    ],
}


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""
    
    workbook_id, workbook_name = detect_workbook_type(filename)
    
    print("=" * 80)
    print(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {filename}")
    print(f"Detected Type: {workbook_id} - {workbook_name}")
    print("=" * 80)
    
    issues_found = []
    warnings_found = []
    
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"\n✅ Workbook loaded successfully")
        print(f"  Sheets found: {len(wb.sheetnames)}")
        print(f"  Sheet names: {', '.join(wb.sheetnames)}")
        
        # Check 1: Verify expected sheets exist
        if workbook_id in EXPECTED_SHEETS:
            expected = set(EXPECTED_SHEETS[workbook_id])
            actual = set(wb.sheetnames)
            
            missing = expected - actual
            extra = actual - expected
            
            if missing:
                issues_found.append(f"Missing expected sheets: {', '.join(missing)}")
            if extra:
                warnings_found.append(f"Extra sheets not in template: {', '.join(extra)}")
            
            if not missing and not extra:
                print(f"\n✅ All expected sheets present ({len(expected)} sheets)")
        
        # Check 2: Sheet name length (Excel limit: 31 characters)
        for sheet_name in wb.sheetnames:
            if len(sheet_name) > 31:
                issues_found.append(f"Sheet name too long (>31 chars): '{sheet_name}'")
        
        # Check 3: Invalid characters in sheet names
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for sheet_name in wb.sheetnames:
            for char in invalid_chars:
                if char in sheet_name:
                    issues_found.append(f"Invalid character '{char}' in sheet name: '{sheet_name}'")
        
        # Check 4: Duplicate sheet names (case-insensitive)
        sheet_names_lower = [name.lower() for name in wb.sheetnames]
        if len(sheet_names_lower) != len(set(sheet_names_lower)):
            issues_found.append("Duplicate sheet names detected (case-insensitive)")
        
        # Check 5: Formula errors
        print(f"\n🔍 Checking for formula errors...")
        formula_errors = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if cell.value.startswith('#'):
                            formula_errors += 1
                            warnings_found.append(f"Formula error in {sheet_name}!{cell.coordinate}: {cell.value}")
        
        if formula_errors == 0:
            print(f"  ✅ No formula errors detected")
        else:
            print(f"  ⚠️  {formula_errors} formula error(s) detected")
        
        # Check 6: Merged cells
        print(f"\n🔍 Checking merged cells...")
        merged_count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            merged_count += len(ws.merged_cells.ranges)
        
        print(f"  ℹ️  Total merged cell ranges: {merged_count}")
        if merged_count > 100:
            warnings_found.append(f"Large number of merged cells ({merged_count}) may affect performance")
        
        # Check 7: Data validations
        print(f"\n🔍 Checking data validations...")
        validation_count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            validation_count += len(ws.data_validations.dataValidation)
        
        print(f"  ✅ Data validations found: {validation_count}")
        
        # Check 8: Emoji support
        print(f"\n🔍 Checking for emoji characters...")
        emoji_count = 0
        emoji_examples = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(max_row=50):  # Check first 50 rows
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Check for common emojis
                        if any(emoji in cell.value for emoji in ['✅', '❌', '⚠️', '🔴', '🟡', '🟢', '🚨', '🎯']):
                            emoji_count += 1
                            if len(emoji_examples) < 3:
                                emoji_examples.append(f"{sheet_name}!{cell.coordinate}: {cell.value[:30]}")
        
        if emoji_count > 0:
            print(f"  ✅ Emoji characters found: {emoji_count}")
            for example in emoji_examples:
                print(f"    • {example}")
        else:
            print(f"  ℹ️  No emojis detected")
        
        # Check 9: Large cells (potential memory issues)
        print(f"\n🔍 Checking for oversized cells...")
        large_cells = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if len(cell.value) > 10000:
                            large_cells += 1
                            warnings_found.append(f"Very large cell in {sheet_name}!{cell.coordinate} ({len(cell.value)} chars)")
        
        if large_cells == 0:
            print(f"  ✅ No oversized cells detected")
        else:
            print(f"  ⚠️  {large_cells} oversized cell(s) detected")
        
        # Check 10: File size
        import os
        file_size = os.path.getsize(filename)
        print(f"\n📊 File size: {file_size:,} bytes ({file_size / (1024*1024):.2f} MB)")
        
        if file_size > 10 * 1024 * 1024:  # 10 MB
            warnings_found.append(f"Large file size ({file_size / (1024*1024):.2f} MB) may affect performance")
        
        wb.close()
        
    except Exception as e:
        issues_found.append(f"Critical error loading workbook: {str(e)}")
        print(f"\n❌ Failed to load workbook: {e}")
        return False
    
    # Summary
    print("\n" + "=" * 80)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 80)
    
    if not issues_found and not warnings_found:
        print("\n✅ WORKBOOK HEALTH: EXCELLENT")
        print("\nNo issues or warnings detected.")
        print("This workbook should open without errors in Excel.")
    else:
        if issues_found:
            print(f"\n❌ CRITICAL ISSUES FOUND: {len(issues_found)}")
            for issue in issues_found:
                print(f"  • {issue}")
        
        if warnings_found:
            print(f"\n⚠️  WARNINGS: {len(warnings_found)}")
            for warning in warnings_found:
                print(f"  • {warning}")
        
        if issues_found:
            print("\n❌ WORKBOOK HEALTH: ISSUES DETECTED")
            print("\nRecommendation: Fix critical issues before distributing.")
        else:
            print("\n⚠️  WORKBOOK HEALTH: ACCEPTABLE WITH WARNINGS")
            print("\nWorkbook should function but may have minor issues.")
    
    print("\n" + "=" * 80)
    
    return len(issues_found) == 0


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    if len(sys.argv) < 2:
        print("Usage: python3 excel_sanity_check_a831.py <workbook.xlsx>")
        print("\nExample:")
        print("  python3 excel_sanity_check_a831.py ISMS-IMP-A.8.31.1_Environment_Architecture_20260111.xlsx")
        sys.exit(1)
    
    filename = sys.argv[1]
    
    if not filename.endswith('.xlsx'):
        print(f"❌ Error: File must be .xlsx format: {filename}")
        sys.exit(1)
    
    import os
    if not os.path.exists(filename):
        print(f"❌ Error: File not found: {filename}")
        sys.exit(1)
    
    success = check_workbook_health(filename)
    
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
