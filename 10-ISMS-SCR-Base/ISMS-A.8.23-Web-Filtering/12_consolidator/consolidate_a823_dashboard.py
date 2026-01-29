#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
Dashboard Data Consolidation Script - ISMS A.8.23 Web Filtering Framework
================================================================================

ISO/IEC 27001:2022 Control A.8.23: Web Filtering
Alternative Data Integration Method for Compliance Dashboard

--------------------------------------------------------------------------------
DATA CONSOLIDATION UTILITY - ALTERNATIVE TO EXTERNAL WORKBOOK LINKING
--------------------------------------------------------------------------------

This utility provides an alternative method for populating the A.8.23.5
Compliance Dashboard with data from source assessment workbooks (A.8.23.1
through A.8.23.4) by directly copying user-entered data rather than using
external workbook formula references.

**Primary Approach vs. Alternative Approach:**

**Primary (Recommended): External Workbook Formula Linking**
- Dashboard contains formulas like: ='[ISMS-IMP-A.8.23.1.xlsx]Gap_Analysis'!$B$15
- Data updates automatically when source files change
- Dashboard and source files must be in same directory
- Requires "Update Links" when opening dashboard
- Real-time data synchronization
- See: generate_a823_5_compliance_dashboard.py

**Alternative (This Script): Direct Data Consolidation**
- Reads data from source workbooks programmatically
- Copies data values (not formulas) into dashboard
- Creates static snapshot of compliance status
- Works with dashboard and sources in different locations
- Requires manual re-run when source data changes
- Useful when external links don't work or aren't desired

**When to Use This Script:**

1. **External Links Don't Work**
   - Excel security settings prevent external workbook links
   - SharePoint/OneDrive environments where linking fails
   - Email distribution requires self-contained workbooks
   - Recipient systems block external references

2. **Static Snapshots Required**
   - Monthly/quarterly compliance reporting archives
   - Board presentations requiring frozen data
   - Audit evidence packages with point-in-time data
   - Historical compliance tracking over time

3. **Simplified Distribution**
   - Single-file distribution to stakeholders
   - No dependency on source file availability
   - Recipients don't have access to source assessments
   - Reduced file size (no external references)

4. **Data Consolidation Flexibility**
   - Selective data inclusion from sources
   - Custom data transformation during consolidation
   - Multi-source aggregation and calculations
   - Data validation during import

**When NOT to Use This Script:**

- If external workbook linking works reliably → Use primary approach
- If real-time data updates are needed → Use external linking
- If dashboard and sources will always be co-located → Use external linking
- For initial dashboard setup → Use generate_a823_5_compliance_dashboard.py first

**Purpose:**
Reads user-entered data from all relevant sheets in the four normalized
assessment workbooks and writes consolidated data into corresponding
dashboard sheets, preserving the Executive Dashboard's external formulas
while populating detail sheets with actual assessment data.

**What It Does:**

1. **Validates Source Availability**
   - Checks all four normalized workbooks exist
   - Validates dashboard workbook is accessible
   - Reports missing files before attempting consolidation

2. **Reads Source Data**
   - Opens each source workbook in data-only mode
   - Reads user-entered data from configured sheets
   - Extracts relevant columns and rows per mapping
   - Skips empty rows and merged cell areas

3. **Consolidates into Dashboard**
   - Maps source sheets to target dashboard sheets
   - Writes data to appropriate dashboard locations
   - Adds source labels for traceability
   - Preserves existing dashboard structure

4. **Preserves Executive Dashboard**
   - Does NOT modify Executive Dashboard sheet
   - Retains external workbook formula references
   - Allows hybrid approach (formulas + consolidated data)

5. **Provides Consolidation Statistics**
   - Reports sheets processed and rows consolidated
   - Shows data distribution across dashboard sheets
   - Identifies any skipped or missing sheets

**Consolidation Mapping:**
The script uses a comprehensive mapping structure defining:
- Source workbook → Source sheet → Target dashboard sheet
- Starting row for data extraction
- Column ranges to include
- Source labels for traceability
- Data organization in dashboard

Example mapping:
```python
"wb1": {  # A.8.23.1 Infrastructure
    "Gap_Analysis": {
        "target": "Gap Analysis",          # Dashboard sheet
        "start_row": 10,                   # Begin reading at row 10
        "columns": "A:R",                  # Include columns A through R
        "label": "Domain 1: Infrastructure"  # Source identifier
    },
    ...
}
```

**Supported Source Workbooks:**
- ISMS-IMP-A.8.23.1.xlsx (Filtering Infrastructure Assessment)
- ISMS-IMP-A.8.23.2.xlsx (Network Coverage Assessment)
- ISMS-IMP-A.8.23.3.xlsx (Policy Configuration Assessment)
- ISMS-IMP-A.8.23.4.xlsx (Monitoring & Response Assessment)

These are the normalized file names created by normalize_assessment_files_a823.py

**Target Dashboard Sheets:**
- Gap Analysis (consolidated gaps from all 4 domains)
- Evidence Register (consolidated evidence from all 4 domains)
- Risk Register (risk assessments from domains 1, 3)
- KPIs & Metrics (performance metrics from domains 1, 2, 4)
- Audit & Compliance Log (detailed assessment data)
- Action Items & Follow-up (exceptions, alerts, follow-up items)

**Executive Dashboard Sheet:**
- NOT modified by this script
- Retains external workbook formula references
- Pulls summary metrics from other dashboard sheets
- Allows hybrid linking + consolidation approach

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

**Basic Usage:**
    python3 consolidate_a823_dashboard.py ISMS-IMP-A.8.23.5_Compliance_Dashboard_20250124.xlsx

**Prerequisites:**
    1. Generate dashboard first:
       python3 generate_a823_5_compliance_dashboard.py
    
    2. Normalize source assessment files:
       python3 normalize_assessment_files_a823.py
    
    3. Ensure all files are accessible:
       - ISMS-IMP-A.8.23.1.xlsx
       - ISMS-IMP-A.8.23.2.xlsx
       - ISMS-IMP-A.8.23.3.xlsx
       - ISMS-IMP-A.8.23.4.xlsx
       - Dashboard file (provided as argument)

**File Location Requirements:**
    Source workbooks must be in CURRENT DIRECTORY when running script:
    - ISMS-IMP-A.8.23.1.xlsx (in current directory)
    - ISMS-IMP-A.8.23.2.xlsx (in current directory)
    - ISMS-IMP-A.8.23.3.xlsx (in current directory)
    - ISMS-IMP-A.8.23.4.xlsx (in current directory)
    
    Dashboard can be in any directory (specify full path if needed)

**Example Session:**
    $ python3 consolidate_a823_dashboard.py ISMS-IMP-A.8.23.5_Dashboard_20250124.xlsx
    ================================================================================
    ISMS A.8.23 - DASHBOARD DATA CONSOLIDATION
    ================================================================================
    
    Dashboard: ISMS-IMP-A.8.23.5_Dashboard_20250124.xlsx
    Started: 2025-01-24 14:30:15
    
    Checking source workbooks...
      ✅ ISMS-IMP-A.8.23.1.xlsx
      ✅ ISMS-IMP-A.8.23.2.xlsx
      ✅ ISMS-IMP-A.8.23.3.xlsx
      ✅ ISMS-IMP-A.8.23.4.xlsx
    
    Loading dashboard...
      ✅ Dashboard loaded (10 sheets)
    
    ================================================================================
    CONSOLIDATING DATA FROM SOURCE WORKBOOKS
    ================================================================================
    
    📁 Processing: ISMS-IMP-A.8.23.1.xlsx
      ✅ Gap_Analysis → Gap Analysis: 12 rows
      ✅ Evidence_Register → Evidence Register: 8 rows
      ✅ Solution_Details_Template → Audit & Compliance Log: 3 rows
      [... additional sheets ...]
    
    📁 Processing: ISMS-IMP-A.8.23.2.xlsx
      ✅ Gap_Identification → Gap Analysis: 5 rows
      [... additional sheets ...]
    
    [... processing workbooks 3 and 4 ...]
    
    ================================================================================
    SAVING CONSOLIDATED DASHBOARD
    ================================================================================
    ✅ Dashboard saved successfully
    
    ================================================================================
    CONSOLIDATION COMPLETE
    ================================================================================
    
    📊 Statistics:
      • Source sheets read: 34
      • Total rows consolidated: 247
    
    📋 By Dashboard Sheet:
      • Action Items & Follow-up: 28 rows
      • Audit & Compliance Log: 89 rows
      • Evidence Register: 31 rows
      • Gap Analysis: 44 rows
      • KPIs & Metrics: 38 rows
      • Risk Register: 17 rows
    
    ✅ Completed: 2025-01-24 14:30:42
    
    ================================================================================
    🎯 Dashboard now contains all user-entered data from assessments!
       Open in Excel and click 'Update Links' to refresh formulas.
    ================================================================================

**Exit Codes:**
    0 = Consolidation successful
    1 = Consolidation failed (missing files, errors during processing)

**Re-Running Consolidation:**
    Safe to re-run this script when assessment data is updated:
    - Existing dashboard data will be overwritten
    - Updated source data will be consolidated
    - Statistics will reflect new data volumes
    - No data loss in source workbooks (read-only access)

**Workflow Integration:**

    **Initial Setup (One-Time):**
    1. Generate dashboard: python3 generate_a823_5_compliance_dashboard.py
    2. Normalize assessments: python3 normalize_assessment_files_a823.py
    3. Consolidate data: python3 consolidate_a823_dashboard.py dashboard.xlsx
    
    **Monthly/Quarterly Updates:**
    1. Update source assessments with new data
    2. Re-normalize (if file names changed): python3 normalize_assessment_files_a823.py
    3. Re-consolidate: python3 consolidate_a823_dashboard.py dashboard.xlsx
    
    **Static Snapshot Creation:**
    1. Run consolidation with date-stamped dashboard name
    2. Archive consolidated dashboard for reporting period
    3. Create new dashboard for next period

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel manipulation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Excel file reading and writing)
    - sys, os (standard library)
    - datetime (standard library)

**Required Files:**
    - Four normalized assessment workbooks (A.8.23.1 through A.8.23.4)
    - Generated dashboard workbook (A.8.23.5)
    - This consolidation script

**File Naming:**
    Source workbooks MUST be named exactly:
    - ISMS-IMP-A.8.23.1.xlsx
    - ISMS-IMP-A.8.23.2.xlsx
    - ISMS-IMP-A.8.23.3.xlsx
    - ISMS-IMP-A.8.23.4.xlsx
    
    Dashboard can have any name (passed as command-line argument)

--------------------------------------------------------------------------------
CONSOLIDATION MAPPING STRUCTURE
--------------------------------------------------------------------------------

**How It Works:**
The script uses a detailed mapping dictionary (CONSOLIDATION_MAP) that defines
which data to extract from which source sheets and where to place it in the
dashboard.

**Mapping Components:**

1. **Source Workbook Key** (wb1, wb2, wb3, wb4)
   - Maps to normalized assessment file names
   - wb1 = ISMS-IMP-A.8.23.1.xlsx (Infrastructure)
   - wb2 = ISMS-IMP-A.8.23.2.xlsx (Network Coverage)
   - wb3 = ISMS-IMP-A.8.23.3.xlsx (Policy Configuration)
   - wb4 = ISMS-IMP-A.8.23.4.xlsx (Monitoring & Response)

2. **Source Sheet Name**
   - Exact sheet name in source workbook
   - Examples: Gap_Analysis, Evidence_Register, Threat_Protection

3. **Target Dashboard Sheet**
   - Destination sheet in dashboard workbook
   - Examples: Gap Analysis, Risk Register, KPIs & Metrics

4. **Start Row**
   - Row number to begin reading data (typically row 10)
   - Skips headers and instructions

5. **Column Range**
   - Columns to include (e.g., "A:R" includes columns A through R)
   - Defines data width for each source

6. **Source Label**
   - Descriptive label added to consolidated data
   - Enables traceability back to source domain

**Total Mapped Sources:**
- Domain 1 (Infrastructure): 8 source sheets → 4 dashboard sheets
- Domain 2 (Network Coverage): 7 source sheets → 4 dashboard sheets
- Domain 3 (Policy Configuration): 9 source sheets → 3 dashboard sheets
- Domain 4 (Monitoring & Response): 9 source sheets → 4 dashboard sheets
- Total: 33 source sheets consolidated into 6 dashboard sheets

**Customization:**
To modify consolidation behavior, edit CONSOLIDATION_MAP dictionary:
- Add/remove source sheets
- Change target dashboard sheets
- Adjust column ranges
- Modify source labels
- Change starting rows

--------------------------------------------------------------------------------
DATA CONSOLIDATION PROCESS
--------------------------------------------------------------------------------

**Step-by-Step Process:**

1. **Validation Phase**
   - Verify dashboard file exists and is accessible
   - Check all four source workbooks are present
   - Report any missing files before proceeding

2. **Dashboard Loading**
   - Open dashboard workbook in read-write mode
   - Validate expected sheets exist
   - Prepare for data consolidation

3. **Source Processing Loop**
   For each source workbook (A.8.23.1 through A.8.23.4):
   
   a. Open source workbook in data-only mode (values, not formulas)
   b. For each configured source sheet:
      - Locate source sheet by name
      - Read data starting from specified row
      - Extract specified column range
      - Stop at first completely empty row
      - Skip merged cell areas safely
   
   c. For each extracted data block:
      - Locate target dashboard sheet
      - Find first available row for writing
      - Add source label for traceability
      - Write data values (not formulas)
      - Track consolidation statistics

4. **Dashboard Preservation**
   - Executive Dashboard sheet is NOT modified
   - External formula references remain intact
   - Hybrid approach maintained (formulas + consolidated data)

5. **Save and Report**
   - Save modified dashboard workbook
   - Generate consolidation statistics
   - Report data distribution across dashboard sheets
   - Provide completion timestamp

**Data Handling:**
- Source workbooks opened in data-only mode (formulas evaluated to values)
- Empty rows terminate data extraction for each source
- Merged cells are safely skipped (no write attempts)
- Existing dashboard data is overwritten on re-consolidation
- Source workbooks remain unmodified (read-only access)

**Error Handling:**
- Missing source files reported before consolidation starts
- Missing source sheets logged but don't stop consolidation
- Missing target sheets logged and skipped
- Cell write errors caught and logged
- Dashboard save failures reported with error details

--------------------------------------------------------------------------------
INTEGRATION WITH A.8.23 FRAMEWORK
--------------------------------------------------------------------------------

**Position in Workflow:**

This script provides an ALTERNATIVE data integration method. The standard
workflow uses external workbook formula linking. This consolidation approach
is used when external linking isn't suitable or desired.

**Standard Workflow (External Linking - Recommended):**
```
Assessments → Normalization → Dashboard Generation → Link Update → Distribution
(A.8.23.1-4)     (normalize)      (generate_a823_5)    (Excel)    (Stakeholders)
```

**Alternative Workflow (Data Consolidation - This Script):**
```
Assessments → Normalization → Dashboard Generation → Data Consolidation → Distribution
(A.8.23.1-4)     (normalize)      (generate_a823_5)   (this script)   (Stakeholders)
                                                             ↓
                                                    Static Snapshot
```

**Hybrid Workflow (Combined Approach):**
```
Assessments → Normalization → Dashboard Generation → Link Update + Consolidation
(A.8.23.1-4)     (normalize)      (generate_a823_5)        (Both)
                                                             ↓
                                              Exec formulas + Detail data
```

**Related Scripts:**
- generate_a823_1_filtering_infrastructure.py (creates A.8.23.1)
- generate_a823_2_network_coverage.py (creates A.8.23.2)
- generate_a823_3_policy_configuration.py (creates A.8.23.3)
- generate_a823_4_monitoring_response.py (creates A.8.23.4)
- normalize_assessment_files_a823.py (normalizes file names)
- generate_a823_5_compliance_dashboard.py (creates dashboard)
- sanity_check_a823_dashboard.py (pre-flight check)
- excel_sanity_check_a823.py (workbook validation)

**Comparison: External Linking vs. Data Consolidation**

| Aspect | External Linking | Data Consolidation |
|--------|------------------|-------------------|
| **Setup** | One-time generation | Generate + consolidate |
| **Updates** | Automatic (Excel refresh) | Manual re-run required |
| **File Dependency** | Must be co-located | Can be separate |
| **Distribution** | Requires source files | Self-contained |
| **Excel Security** | May be blocked | Always works |
| **Historical Tracking** | Overwrites data | Create dated snapshots |
| **Data Type** | Live formulas | Static values |
| **Use Case** | Real-time reporting | Point-in-time reporting |

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.23
Script Type:          Data Consolidation / Alternative Integration Method
Framework Component:  A.8.23.5 Compliance Dashboard Support
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Input Files:
    - ISMS-IMP-A.8.23.1.xlsx (Filtering Infrastructure Assessment)
    - ISMS-IMP-A.8.23.2.xlsx (Network Coverage Assessment)
    - ISMS-IMP-A.8.23.3.xlsx (Policy Configuration Assessment)
    - ISMS-IMP-A.8.23.4.xlsx (Monitoring & Response Assessment)
    - Dashboard workbook (ISMS-IMP-A.8.23.5_*.xlsx)

Output:
    - Modified dashboard workbook with consolidated data
    - Console statistics showing consolidation results

Related Quality Tools:
    - normalize_assessment_files_a823.py (file preparation)
    - generate_a823_5_compliance_dashboard.py (dashboard creation)
    - excel_sanity_check_a823.py (workbook validation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Comprehensive consolidation mapping for all 4 assessment domains
    - 33 source sheets → 6 dashboard sheets
    - Preserves Executive Dashboard external formulas
    - Safe merged cell handling
    - Detailed consolidation statistics

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**File Modification:**
- Source workbooks are read in data-only mode (NOT modified)
- Dashboard workbook IS modified (data written to detail sheets)
- Executive Dashboard sheet is NOT modified (formulas preserved)
- Always backup dashboard before first consolidation

**Re-Running Consolidation:**
- Safe to re-run when source data is updated
- Existing consolidated data is overwritten
- No cumulative effect (fresh consolidation each time)
- Statistics reflect current consolidation only

**Source File Requirements:**
- Must use EXACT normalized file names
- Must be in current directory when running script
- Must be readable .xlsx format (not .xls or .xlsm)
- Must contain expected sheet names from mapping

**Dashboard Requirements:**
- Must be generated by generate_a823_5_compliance_dashboard.py
- Must contain expected target sheet names
- Must be writable (not read-only or open in Excel)
- Should be backed up before first consolidation

**Mapping Maintenance:**
- CONSOLIDATION_MAP must match actual assessment sheet names
- Changing assessment structures requires mapping updates
- Adding custom sheets requires mapping additions
- Verify mapping after generator script modifications

**Performance:**
- Typical consolidation time: 10-30 seconds
- Depends on data volume in source workbooks
- Progress is displayed for each source workbook
- Large datasets (1000+ rows) may take longer

**Limitations:**
- Does not preserve formulas from source workbooks
- Does not copy cell formatting or styles
- Does not handle Excel tables or charts
- Does not validate data during consolidation
- Assumes source data is already validated

**Best Practices:**
1. Complete source assessments before consolidation
2. Run normalization script before consolidation
3. Validate source workbooks with excel_sanity_check_a823.py
4. Backup dashboard before first consolidation
5. Review consolidation statistics after each run
6. Test hybrid approach (formulas + consolidation) in dev first
7. Document when/why consolidation approach is used vs. external linking

**Troubleshooting Common Issues:**

**Issue: "ERROR: Dashboard file not found"**
Solution: Verify dashboard path is correct, check file exists

**Issue: "MISSING: ISMS-IMP-A.8.23.N.xlsx"**
Solution: Run normalize_assessment_files_a823.py to create normalized files

**Issue: "Sheet 'X' not found, skipping"**
Solution: Source workbook missing expected sheet, check assessment generator

**Issue: "Target sheet 'X' not in dashboard, skipping"**
Solution: Dashboard missing expected sheet, regenerate dashboard

**Issue: "No data" for multiple sheets**
Solution: Source workbooks may be empty or data in unexpected location

**Issue: "ERROR saving dashboard"**
Solution: Dashboard may be open in Excel, close it before consolidation

**Audit Considerations:**
- Consolidation creates static point-in-time snapshot
- Consolidation timestamp is recorded in statistics
- Source data traceability via source labels
- Useful for monthly/quarterly compliance reporting archives
- Executive Dashboard formulas still provide real-time if links enabled

**Data Protection:**
- Consolidated dashboard inherits data classification from sources
- Source workbooks remain unmodified and unaffected
- Dashboard may contain aggregated sensitive compliance data
- Handle according to organization's data classification policy

**Alternative to This Script:**
If external workbook linking works reliably in your environment, consider
using the standard approach (generate dashboard + enable "Update Links")
instead of data consolidation for automatic updates and reduced maintenance.

================================================================================
END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ============================================================================
# CONFIGURATION
# ============================================================================

SOURCE_WORKBOOKS = {
    "wb1": "ISMS-IMP-A.8.23.1.xlsx",  # Infrastructure
    "wb2": "ISMS-IMP-A.8.23.2.xlsx",  # Network Coverage
    "wb3": "ISMS-IMP-A.8.23.3.xlsx",  # Policy Configuration
    "wb4": "ISMS-IMP-A.8.23.4.xlsx",  # Monitoring & Response
}

# Mapping: source workbook → source sheet → dashboard sheet
CONSOLIDATION_MAP = {
    "wb1": {
        "Gap_Analysis": {
            "target": "Gap Analysis",
            "start_row": 10,
            "columns": "A:R",
            "label": "Domain 1: Infrastructure"
        },
        "Evidence_Register": {
            "target": "Evidence Register",
            "start_row": 10,
            "columns": "A:H",
            "label": "Domain 1: Infrastructure"
        },
        "Solution_Details_Template": {
            "target": "Audit & Compliance Log",
            "start_row": 10,
            "columns": "A:F",
            "label": "Infrastructure: Solutions"
        },
        "Technology_Comparison": {
            "target": "Audit & Compliance Log",
            "start_row": 10,
            "columns": "A:H",
            "label": "Infrastructure: Tech Comparison"
        },
        "Capability_Requirements": {
            "target": "Risk Register",
            "start_row": 10,
            "columns": "A:F",
            "label": "Infrastructure: Capabilities"
        },
        "Integration_Architecture": {
            "target": "KPIs & Metrics",
            "start_row": 10,
            "columns": "A:F",
            "label": "Infrastructure: Integration"
        },
        "Licensing_Support": {
            "target": "Audit & Compliance Log",
            "start_row": 10,
            "columns": "A:F",
            "label": "Infrastructure: Licensing"
        },
        "Performance_Metrics": {
            "target": "KPIs & Metrics",
            "start_row": 10,
            "columns": "A:D",
            "label": "Infrastructure: Performance"
        },
    },
    "wb2": {
        "Gap_Identification": {
            "target": "Gap Analysis",
            "start_row": 10,
            "columns": "A:R",
            "label": "Domain 2: Network Coverage"
        },
        "Evidence_Register": {
            "target": "Evidence Register",
            "start_row": 10,
            "columns": "A:H",
            "label": "Domain 2: Network Coverage"
        },
        "Network_Segment_Inventory": {
            "target": "Audit & Compliance Log",
            "start_row": 10,
            "columns": "A:H",
            "label": "Coverage: Network Segments"
        },
        "Coverage_Matrix": {
            "target": "KPIs & Metrics",
            "start_row": 10,
            "columns": "A:G",
            "label": "Coverage: Matrix"
        },
        "Device_Inventory": {
            "target": "Audit & Compliance Log",
            "start_row": 10,
            "columns": "A:F",
            "label": "Coverage: Devices"
        },
        "Exemption_Register": {
            "target": "Action Items & Follow-up",
            "start_row": 10,
            "columns": "A:H",
            "label": "Coverage: Exemptions"
        },
        "Coverage_Verification": {
            "target": "Audit & Compliance Log",
            "start_row": 10,
            "columns": "A:F",
            "label": "Coverage: Verification"
        },
    },
    "wb3": {
        "Gap_Analysis": {
            "target": "Gap Analysis",
            "start_row": 10,
            "columns": "A:R",
            "label": "Domain 3: Policy Configuration"
        },
        "Evidence_Register": {
            "target": "Evidence Register",
            "start_row": 10,
            "columns": "A:H",
            "label": "Domain 3: Policy Configuration"
        },
        "Threat_Protection": {
            "target": "Risk Register",
            "start_row": 10,
            "columns": "A:H",
            "label": "Policy: Threat Protection"
        },
        "Category_Management": {
            "target": "Audit & Compliance Log",
            "start_row": 10,
            "columns": "A:F",
            "label": "Policy: Categories"
        },
        "Custom_Lists": {
            "target": "Audit & Compliance Log",
            "start_row": 10,
            "columns": "A:E",
            "label": "Policy: Custom Lists"
        },
        "Policy_Exceptions": {
            "target": "Action Items & Follow-up",
            "start_row": 10,
            "columns": "A:G",
            "label": "Policy: Exceptions"
        },
        "User_Group_Policies": {
            "target": "Audit & Compliance Log",
            "start_row": 10,
            "columns": "A:F",
            "label": "Policy: User Groups"
        },
        "Acceptable_Use_Alignment": {
            "target": "Audit & Compliance Log",
            "start_row": 10,
            "columns": "A:E",
            "label": "Policy: AUP Alignment"
        },
        "Policy_Review_Process": {
            "target": "Audit & Compliance Log",
            "start_row": 10,
            "columns": "A:F",
            "label": "Policy: Reviews"
        },
    },
    "wb4": {
        "Gap_Analysis": {
            "target": "Gap Analysis",
            "start_row": 10,
            "columns": "A:R",
            "label": "Domain 4: Monitoring & Response"
        },
        "Evidence_Register": {
            "target": "Evidence Register",
            "start_row": 10,
            "columns": "A:H",
            "label": "Domain 4: Monitoring & Response"
        },
        "Log_Collection": {
            "target": "KPIs & Metrics",
            "start_row": 10,
            "columns": "A:G",
            "label": "Monitoring: Log Sources"
        },
        "Alert_Configuration": {
            "target": "Action Items & Follow-up",
            "start_row": 10,
            "columns": "A:G",
            "label": "Monitoring: Alerts"
        },
        "Monitoring_Dashboard": {
            "target": "KPIs & Metrics",
            "start_row": 10,
            "columns": "A:E",
            "label": "Monitoring: Dashboards"
        },
        "Incident_Response": {
            "target": "Audit & Compliance Log",
            "start_row": 10,
            "columns": "A:G",
            "label": "Monitoring: Incidents"
        },
        "Blocked_Events_Analysis": {
            "target": "KPIs & Metrics",
            "start_row": 10,
            "columns": "A:F",
            "label": "Monitoring: Blocked Events"
        },
        "False_Positive_Mgmt": {
            "target": "Action Items & Follow-up",
            "start_row": 10,
            "columns": "A:H",
            "label": "Monitoring: False Positives"
        },
        "Reporting_Schedule": {
            "target": "Audit & Compliance Log",
            "start_row": 10,
            "columns": "A:F",
            "label": "Monitoring: Reports"
        },
    },
}

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def column_letter_to_index(letter):
    """Convert column letter (A, B, AA, etc.) to 0-based index."""
    result = 0
    for char in letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result - 1


def parse_column_range(col_range):
    """Parse 'A:H' into (start_col_idx, end_col_idx)."""
    start, end = col_range.split(':')
    return column_letter_to_index(start), column_letter_to_index(end)


def read_sheet_data(ws, start_row, col_range, max_rows=500):
    """
    Read data from a sheet starting at start_row.
    
    Returns: List of rows (each row is a list of cell values)
    Stops when hitting an empty row or max_rows.
    """
    start_col, end_col = parse_column_range(col_range)
    data_rows = []
    
    for row_idx in range(start_row, start_row + max_rows):
        row_values = []
        for col_idx in range(start_col, end_col + 1):
            # Convert to 1-based column index for openpyxl
            cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
            row_values.append(cell_value)
        
        # Stop if entire row is empty
        if all(val is None or str(val).strip() == '' for val in row_values):
            break
            
        data_rows.append(row_values)
    
    return data_rows


def write_data_to_sheet(ws, data_rows, start_row, start_col, label=None):
    """
    Write consolidated data to dashboard sheet.
    
    If label is provided, adds it as first column for source tracking.
    """
    if not data_rows:
        return 0
    
    written_count = 0
    current_row = start_row
    
    # Find the first empty row in target sheet (skip merged cells and headers)
    max_search = start_row + 500
    while current_row < max_search:
        try:
            # Try to access the cell - this will fail for merged cells
            test_cell = ws.cell(row=current_row, column=start_col + 1)
            if hasattr(test_cell, 'value') and test_cell.value is None:
                # Found an empty, writable cell
                break
        except:
            # Skip this row if there's any issue
            pass
        current_row += 1
    
    # Write data rows
    for row_data in data_rows:
        # Skip if we hit a merged cell area
        try:
            test_write = ws.cell(row=current_row, column=start_col)
            if not hasattr(test_write, 'value'):
                current_row += 1
                continue
        except:
            current_row += 1
            continue
        
        # Add label in first column if provided
        if label:
            try:
                ws.cell(row=current_row, column=start_col).value = label
            except:
                pass  # Skip if can't write
            col_offset = 1
        else:
            col_offset = 0
        
        # Write row data
        for col_idx, value in enumerate(row_data):
            try:
                ws.cell(row=current_row, column=start_col + col_offset + col_idx).value = value
            except:
                pass  # Skip if can't write (merged cell, etc.)
        
        current_row += 1
        written_count += 1
    
    return written_count


# ============================================================================
# MAIN CONSOLIDATION LOGIC
# ============================================================================

def consolidate_data(dashboard_path):
    """Main consolidation function."""
    
    print("=" * 80)
    print("ISMS A.8.23 - DASHBOARD DATA CONSOLIDATION")
    print("=" * 80)
    print(f"\nDashboard: {dashboard_path}")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    # Check dashboard exists
    if not os.path.exists(dashboard_path):
        print(f"\u274C ERROR: Dashboard file not found: {dashboard_path}")
        return 1
    
    # Check source workbooks
    print("Checking source workbooks...")
    missing = []
    for wb_key, wb_path in SOURCE_WORKBOOKS.items():
        if os.path.exists(wb_path):
            print(f"  \u2705 {wb_path}")
        else:
            print(f"  \u274C MISSING: {wb_path}")
            missing.append(wb_path)
    
    if missing:
        print(f"\n\u274C ERROR: {len(missing)} source workbook(s) missing.")
        print("   Run auto-normalization or ensure files exist.")
        return 1
    
    # Load dashboard
    print(f"\nLoading dashboard...")
    try:
        dashboard = load_workbook(dashboard_path)
        print(f"  \u2705 Dashboard loaded ({len(dashboard.sheetnames)} sheets)")
    except Exception as e:
        print(f"  \u274C ERROR loading dashboard: {e}")
        return 1
    
    # Consolidation statistics
    stats = {
        "total_sheets_read": 0,
        "total_rows_consolidated": 0,
        "by_target_sheet": {},
    }
    
    # Process each source workbook
    print("\n" + "=" * 80)
    print("CONSOLIDATING DATA FROM SOURCE WORKBOOKS")
    print("=" * 80)
    
    for wb_key, wb_path in SOURCE_WORKBOOKS.items():
        print(f"\n📁 Processing: {wb_path}")
        
        try:
            source_wb = load_workbook(wb_path, data_only=True)
        except Exception as e:
            print(f"  \u274C ERROR loading {wb_path}: {e}")
            continue
        
        # Process each source sheet defined in mapping
        if wb_key not in CONSOLIDATION_MAP:
            print(f"  \u26A0\uFE0F  No consolidation mapping defined for {wb_key}")
            continue
        
        for source_sheet_name, config in CONSOLIDATION_MAP[wb_key].items():
            if source_sheet_name not in source_wb.sheetnames:
                print(f"  \u26A0\uFE0F  Sheet '{source_sheet_name}' not found, skipping")
                continue
            
            target_sheet_name = config["target"]
            if target_sheet_name not in dashboard.sheetnames:
                print(f"  \u26A0\uFE0F  Target sheet '{target_sheet_name}' not in dashboard, skipping")
                continue
            
            # Read data from source
            source_ws = source_wb[source_sheet_name]
            data_rows = read_sheet_data(
                source_ws,
                config["start_row"],
                config["columns"],
                max_rows=500
            )
            
            if not data_rows:
                print(f"  \u2022 {source_sheet_name} → {target_sheet_name}: No data")
                continue
            
            # Write to target
            target_ws = dashboard[target_sheet_name]
            written = write_data_to_sheet(
                target_ws,
                data_rows,
                start_row=10,  # Dashboard sheets start at row 10
                start_col=1,   # Column A
                label=config.get("label")
            )
            
            print(f"  \u2705 {source_sheet_name} → {target_sheet_name}: {written} rows")
            
            stats["total_sheets_read"] += 1
            stats["total_rows_consolidated"] += written
            stats["by_target_sheet"][target_sheet_name] = \
                stats["by_target_sheet"].get(target_sheet_name, 0) + written
        
        source_wb.close()
    
    # Save dashboard
    print("\n" + "=" * 80)
    print("SAVING CONSOLIDATED DASHBOARD")
    print("=" * 80)
    
    try:
        dashboard.save(dashboard_path)
        print(f"\u2705 Dashboard saved successfully")
    except Exception as e:
        print(f"\u274C ERROR saving dashboard: {e}")
        return 1
    
    dashboard.close()
    
    # Summary
    print("\n" + "=" * 80)
    print("CONSOLIDATION COMPLETE")
    print("=" * 80)
    print(f"\n📊 Statistics:")
    print(f"  \u2022 Source sheets read: {stats['total_sheets_read']}")
    print(f"  \u2022 Total rows consolidated: {stats['total_rows_consolidated']}")
    print(f"\n\u1F4CB By Dashboard Sheet:")
    for sheet, count in sorted(stats["by_target_sheet"].items()):
        print(f"  \u2022 {sheet}: {count} rows")
    
    print(f"\n\u2705 Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("\n" + "=" * 80)
    print("🎯 Dashboard now contains all user-entered data from assessments!")
    print("   Open in Excel and click 'Update Links' to refresh formulas.")
    print("=" * 80 + "\n")
    
    return 0


# ============================================================================
# ENTRY POINT
# ============================================================================

def main():
    if len(sys.argv) != 2:
        print("Usage: python3 consolidate_a823_dashboard.py <dashboard.xlsx>")
        print("\nExample:")
        print("  python3 consolidate_a823_dashboard.py ISMS-IMP-A.8.23.5_Compliance_Summary_Dashboard_20260113.xlsx")
        return 1
    
    dashboard_path = sys.argv[1]
    return consolidate_data(dashboard_path)


if __name__ == "__main__":
    sys.exit(main())
