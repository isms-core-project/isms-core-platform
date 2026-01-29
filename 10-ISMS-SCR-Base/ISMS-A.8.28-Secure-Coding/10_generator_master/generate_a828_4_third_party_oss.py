#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.28.4 - Third-Party & Open Source Software Assessment
================================================================================

ISO/IEC 27001:2022 Control A.8.28: Secure Coding
Assessment Domain 4 of 4: Third-Party Code and Open Source Component Security

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific dependency management practices, vendor evaluation
criteria, and open source governance policies.

Key customization areas:
1. Dependency management tools and processes (match your package managers)
2. Open source approval criteria and risk thresholds (adapt to risk tolerance)
3. Vendor security assessment requirements (specific to procurement policies)
4. License compliance requirements (based on legal/IP policies)
5. Vulnerability remediation SLAs (aligned with risk management framework)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.28 Secure Coding Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
third-party code security and open source component risk management across all
applications and development projects.

**Purpose:**
Enables systematic assessment of third-party and open source software security
against ISO 27001:2022 Control A.8.28 requirements, supporting evidence-based
validation of supply chain security and dependency risk management practices.

**Assessment Scope:**
- Software Composition Analysis (SCA) tool implementation and coverage
- Open source component inventory and version tracking
- Vulnerability scanning for dependencies and transitive dependencies
- License compliance tracking and approval workflows
- Third-party vendor security assessment processes
- Commercial/COTS software security evaluation
- Dependency update policies and patch management
- End-of-life (EOL) and unmaintained component identification
- Software Bill of Materials (SBOM) generation and maintenance
- Supply chain attack risk mitigation (dependency confusion, typosquatting)
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and OSS security standards
2. SCA Tool Coverage - Software Composition Analysis implementation assessment
3. Component Inventory - Open source and third-party component catalog
4. Vulnerability Management - Dependency vulnerability tracking and remediation
5. License Compliance - OSS license tracking and approval workflow
6. Vendor Assessment - Third-party vendor security evaluation process
7. Update Policy - Dependency update cadence and patch management
8. EOL Components - End-of-life and unmaintained dependency identification
9. SBOM Management - Software Bill of Materials generation and accuracy
10. Supply Chain Risk - Dependency confusion and supply chain attack mitigation
11. Gap Analysis - Inadequate controls or high-risk dependencies and remediation
12. Evidence Register - Audit evidence tracking and documentation
13. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with dependency type and risk level dropdown lists
- Conditional formatting for vulnerability severity and remediation status
- Automated gap identification for high-risk or EOL dependencies
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with SCA tool outputs and vulnerability databases

**Integration:**
This assessment feeds into the A.8.28.5 Compliance Dashboard, which
consolidates data from all four assessment domains for executive
oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl >= 3.0.0 (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a828_4_third_party_oss.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a828_4_third_party_oss.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a828_4_third_party_oss.py --date 20250124

Output:
    File: ISMS_A_8_28_4_Third_Party_OSS_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize OSS governance criteria to match your risk profile
    2. Inventory all applications and their dependency manifests
    3. Complete assessments for each application or microservice
    4. Validate SCA tool coverage and vulnerability detection accuracy
    5. Review EOL component list and plan migration/updates
    6. Conduct gap analysis for high-risk dependencies or missing controls
    7. Define remediation actions with timelines and ownership
    8. Collect and link audit evidence (SCA reports, SBOMs, vendor assessments)
    9. Obtain stakeholder approvals (AppSec, Legal/Compliance, Procurement)
    10. Feed results into A.8.28.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.28
Assessment Domain:    4 of 4 (Third-Party & OSS)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization ISMS Team]
Date:                 DD.MM.YYYY
Last Modified:        DD.MM.YYYY
Python Version:       3.8+
License:              [Organization License/Terms]

Related Documents:
    - ISMS-POL-A.8.28: Secure Coding Policy (Governance)
    - ISMS-IMP-A.8.28.4: Third-Party & OSS Implementation Guide
    - ISMS-IMP-A.8.28.1: SDLC Integration Assessment (Domain 1)
    - ISMS-IMP-A.8.28.2: Standards & Tools Assessment (Domain 2)
    - ISMS-IMP-A.8.28.3: Code Review & Testing Assessment (Domain 3)
    - ISMS-IMP-A.8.28.5: Compliance Dashboard (Consolidation)

Related Scripts:
    - generate_a828_1_sdlc_assessment.py
    - generate_a828_2_standards_tools.py
    - generate_a828_3_code_review_testing.py
    - generate_a828_5_compliance_dashboard.py
    - normalize_assessment_files_a828.py

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - DD.MM.YYYY
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.28.4 specification
    - Supports comprehensive third-party and OSS security evaluation
    - Integrated with A.8.28.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Supply Chain Security Reality:**
Third-party and open source dependencies are the #1 attack vector in modern
software. SolarWinds, Log4Shell, event-stream npm package - all supply chain
attacks. This assessment domain is not optional or "nice to have" - it's
critical infrastructure security.

**Open Source ≠ Insecure:**
Don't treat all OSS as risky. Well-maintained OSS projects (Linux, OpenSSL,
React, etc.) often have better security than commercial alternatives due to
transparency and community review. Focus on maintenance activity, not just
presence of CVEs.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of SCA tool coverage, SBOM accuracy, and
vulnerability remediation for high/critical findings.

**Data Protection:**
Assessment workbooks contain sensitive supply chain information including:
- Complete dependency manifests and software architecture
- Vulnerability details and exploitability assessments
- Vendor security evaluation results and risk ratings
- License compliance issues and legal considerations

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Weekly: Check for new high/critical vulnerabilities in dependencies
- Monthly: Review EOL component list and update dependencies
- Quarterly: Complete SCA tool coverage and accuracy validation
- Annually: Full reassessment of all applications and dependency policies
- Ad-hoc: When major supply chain incidents occur (Log4Shell-scale events)

**Quality Assurance:**
Have application security SMEs, software architects, and legal/compliance
teams validate assessments before using results for compliance reporting or
remediation decisions. Legal review is essential for license compliance.

**Regulatory Alignment:**
Ensure third-party software security aligns with applicable requirements:
- Payment processing: PCI DSS third-party service provider requirements
- Healthcare: HIPAA Business Associate Agreements for software vendors
- Finance: Regional banking third-party risk management requirements
- Government: Software supply chain security (SBOM requirements, EO 14028)

Customize assessment criteria to include regulatory-specific requirements.

**Integration with Dependency Tools:**
Where possible, integrate assessment data with:
- SCA tools (Snyk, WhiteSource, Black Duck, Dependabot, etc.)
- Package managers (npm, Maven, pip, NuGet, Gradle, etc.)
- Vulnerability databases (NVD, GitHub Advisory, OSV, etc.)
- SBOM generation tools (Syft, CycloneDX, SPDX tools)
- License compliance platforms (FOSSA, FOSSology, etc.)

Automated tool integration reduces manual effort and improves accuracy.

**Don't Fool Yourself - Feynman Principle:**
Just because you have an SCA tool doesn't mean you're secure. Key questions:
- Do you actually know all your dependencies? (Including transitive ones?)
- When a critical vulnerability is announced, can you identify affected apps
  within hours, not weeks?
- Do you have a process to update dependencies, or just detect vulnerabilities?
- Are developers empowered to update dependencies, or blocked by bureaucracy?
- Can you generate an accurate SBOM for any application on demand?

If you can't answer "yes" to these, you have dependency visibility problems,
not just dependency security problems. Fix visibility first.

**Dependency Update Philosophy:**
Staying current with dependencies is cheaper than emergency patching during
active exploitation. Regular dependency updates (monthly/quarterly) prevent
technical debt accumulation. Zero-day vulnerabilities in outdated dependencies
are especially dangerous because exploits are often public.

Automate dependency updates where possible (Dependabot, Renovate, etc.) and
treat dependency updates as normal maintenance, not exceptional events.

================================================================================
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •

from datetime import datetime
import sys

# ============================================================================
# SECTION 1: WORKBOOK INITIALIZATION AND STYLE DEFINITIONS
# ============================================================================

def create_workbook():
    """Initialize workbook with proper metadata."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Set workbook properties
    wb.properties.title = "ISMS Control A.8.28.4 - Third-Party & OSS Assessment"
    wb.properties.subject = "Supply Chain Security Assessment"
    wb.properties.creator = "ISMS Assessment Tool"
    wb.properties.description = "Assessment workbook for third-party dependencies and open source software management"
    
    return wb


def get_style_definitions():
    """
    Define standard style elements for consistent formatting.
    
    Color Palette (Organizational Standard):
        - Headers: Dark blue (#003366) with white text
        - Subheaders: Medium blue (#4472C4) with white text
        - Section headers: Green (#70AD47)
        - Column headers: Light gray (#D9D9D9)
        - Input cells: Pale yellow (#FFFFCC)
        - Status colors: Green (#C6EFCE), Yellow (#FFEB9C), Red (#FFC7CE), Gray (#E7E6E6)
        - Priority colors: Dark red (#C00000), Light red (#FF6666), Yellow (#FFEB9C), Light green (#C6EFCE)
    """
    styles = {
        # Header styles
        'header_font': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
        'header_fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        
        'subheader_font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
        'subheader_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        
        'section_font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'section_fill': PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid'),
        
        # Column header styles
        'column_header_font': Font(name='Calibri', size=10, bold=True),
        'column_header_fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        
        # Content styles
        'normal_font': Font(name='Calibri', size=10),
        'bold_font': Font(name='Calibri', size=10, bold=True),
        'italic_font': Font(name='Calibri', size=10, italic=True),
        
        # Input cell style
        'input_fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
        
        # Status color fills
        'status_green': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'status_yellow': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'status_red': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'status_gray': PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'),
        
        # Priority color fills
        'priority_critical_fill': PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'),
        'priority_critical_font': Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
        'priority_high_fill': PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid'),
        'priority_medium_fill': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'priority_low_fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        
        # Alignment styles
        'align_left': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'align_center': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'align_right': Alignment(horizontal='right', vertical='top', wrap_text=True),
        
        # Border styles
        'thin_border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'thick_border': Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        ),
    }
    
    return styles


def get_column_widths():
    """
    Define standard column widths for assessment sheets.
    
    Returns:
        dict: Column letter -> width mapping
    """
    return {
        'A': 8,   # ID column
        'B': 55,  # Requirement description
        'C': 22,  # Implementation Status dropdown
        'D': 35,  # Evidence Reference
        'E': 40,  # Comments/Notes
        'F': 12,  # Compliance indicator (auto-calculated)
    }


# ============================================================================
# SECTION 2: DATA VALIDATION DEFINITIONS
# ============================================================================

def create_data_validations():
    """
    Create reusable data validation objects (dropdowns) for the workbook.
    
    Returns:
        dict: Validation name -> DataValidation object mapping
    """
    validations = {}
    
    # Implementation Status dropdown (most common)
    validations['implementation_status'] = DataValidation(
        type="list",
        formula1='"Implemented,Partially Implemented,Not Implemented,N/A"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Status",
        error="Please select a valid implementation status from the dropdown."
    )
    validations['implementation_status'].prompt = "Select implementation status"
    validations['implementation_status'].promptTitle = "Implementation Status"
    
    # Yes/No dropdown
    validations['yes_no'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    validations['yes_no'].prompt = "Select Yes or No"
    validations['yes_no'].promptTitle = "Yes/No"
    
    # Yes/No/N/A dropdown
    validations['yes_no_na'] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=False
    )
    validations['yes_no_na'].prompt = "Select Yes, No, or N/A"
    validations['yes_no_na'].promptTitle = "Yes/No/N/A"
    
    # Priority dropdown
    validations['priority'] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    validations['priority'].prompt = "Select priority level"
    validations['priority'].promptTitle = "Priority"
    
    # Gap Status dropdown (for Gap Analysis sheet)
    validations['gap_status'] = DataValidation(
        type="list",
        formula1='"Open,In Progress,Resolved,Closed,Deferred"',
        allow_blank=False
    )
    validations['gap_status'].prompt = "Select gap status"
    validations['gap_status'].promptTitle = "Gap Status"
    
    # Evidence Type dropdown
    validations['evidence_type'] = DataValidation(
        type="list",
        formula1='"Document,Screenshot,Report,Configuration,URL,Recording,Database Query,Log File,Other"',
        allow_blank=False
    )
    validations['evidence_type'].prompt = "Select evidence type"
    validations['evidence_type'].promptTitle = "Evidence Type"
    
    # Verification Status dropdown (for Evidence Register)
    validations['verification_status'] = DataValidation(
        type="list",
        formula1='"Pending,Verified,Rejected"',
        allow_blank=False
    )
    validations['verification_status'].prompt = "Select verification status"
    validations['verification_status'].promptTitle = "Verification Status"
    
    # Approval Decision dropdown (for Approval Sign-Off)
    validations['approval_decision'] = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Pending Review"',
        allow_blank=False
    )
    validations['approval_decision'].prompt = "Select approval decision"
    validations['approval_decision'].promptTitle = "Approval Decision"
    
    # Domain dropdown (for Gap Analysis)
    validations['domain'] = DataValidation(
        type="list",
        formula1='"Vendor Security Assessment,OSS Management,Dependency Vulnerability Mgmt,Third-Party Code Review,License Compliance"',
        allow_blank=False
    )
    validations['domain'].prompt = "Select assessment domain"
    validations['domain'].promptTitle = "Domain"
    
    return validations


def apply_validation_to_range(ws, validation, start_row, end_row, column):
    """
    Apply a data validation to a range of cells.
    
    Args:
        ws: Worksheet object
        validation: DataValidation object
        start_row: Starting row number
        end_row: Ending row number
        column: Column letter (e.g., 'C')
    """
    ws.add_data_validation(validation)
    for row in range(start_row, end_row + 1):
        cell_ref = f"{column}{row}"
        validation.add(cell_ref)


def apply_standard_formatting(ws, start_row, end_row, styles, column_widths):
    """
    Apply standard formatting to assessment domain rows.
    
    Args:
        ws: Worksheet object
        start_row: First data row
        end_row: Last data row
        styles: Style definitions dictionary
        column_widths: Column width definitions
    """
    # Set column widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Apply borders and alignment to all cells
    for row in range(start_row, end_row + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f'{col}{row}']
            cell.border = styles['thin_border']
            cell.font = styles['normal_font']
            
            # Column-specific alignment
            if col == 'A':
                cell.alignment = Alignment(horizontal='center', vertical='top')
            elif col == 'F':
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = styles['align_left']
    
    # Highlight input columns (C, D, E) with pale yellow
    for row in range(start_row, end_row + 1):
        for col in ['C', 'D', 'E']:
            ws[f'{col}{row}'].fill = styles['input_fill']


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def set_header_row(ws, row_num, headers, styles):
    """
    Create a formatted header row.
    
    Args:
        ws: Worksheet object
        row_num: Row number for headers
        headers: List of header texts
        styles: Style definitions dictionary
    """
    for idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=idx)
        cell.value = header_text
        cell.font = styles['column_header_font']
        cell.fill = styles['column_header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['thin_border']


def freeze_header_rows(ws, row_num):
    """
    Freeze rows above the specified row number.
    
    Args:
        ws: Worksheet object
        row_num: Row number to freeze above (e.g., 4 freezes rows 1-3)
    """
    ws.freeze_panes = f'A{row_num}'


def add_title_section(ws, row_num, title, styles):
    """
    Add a formatted title section spanning multiple columns.
    
    Args:
        ws: Worksheet object
        row_num: Row number for title
        title: Title text
        styles: Style definitions dictionary
    
    Returns:
        int: Next available row number
    """
    ws.merge_cells(f'A{row_num}:F{row_num}')
    cell = ws[f'A{row_num}']
    cell.value = title
    cell.font = styles['header_font']
    cell.fill = styles['header_fill']
    cell.alignment = styles['align_center']
    cell.border = styles['thick_border']
    
    return row_num + 1


def add_metadata_row(ws, row_num, label, value, styles):
    """
    Add a metadata row (e.g., "Date: 07.01.2025").
    
    Args:
        ws: Worksheet object
        row_num: Row number
        label: Label text (e.g., "Assessment Date:")
        value: Value text
        styles: Style definitions dictionary
    
    Returns:
        int: Next available row number
    """
    ws[f'A{row_num}'] = label
    ws[f'A{row_num}'].font = styles['bold_font']
    ws[f'A{row_num}'].alignment = styles['align_left']
    
    ws.merge_cells(f'B{row_num}:F{row_num}')
    ws[f'B{row_num}'] = value
    ws[f'B{row_num}'].font = styles['normal_font']
    ws[f'B{row_num}'].alignment = styles['align_left']
    
    return row_num + 1

# ============================================================================
# SECTION 3: ASSESSMENT DOMAIN DATA DEFINITIONS
# ============================================================================

def get_vendor_security_requirements():
    """
    Domain 1: Vendor Security Assessment (18 requirements)
    
    Focus: Vendor due diligence, contracts, access controls, risk management
    
    Returns:
        list: List of tuples (ID, Requirement, Evidence Examples)
    """
    requirements = [
        (
            "V-01",
            "Vendor security assessment policy exists and is approved",
            "Policy document, approval signatures, policy publication records"
        ),
        (
            "V-02",
            "Security questionnaire sent to all vendors before onboarding",
            "Questionnaire templates, completed questionnaires, vendor onboarding records"
        ),
        (
            "V-03",
            "Vendor risk classification framework implemented (High/Medium/Low tiers)",
            "Risk classification matrix, vendor tier assignments, risk scoring methodology"
        ),
        (
            "V-04",
            "Contract security requirements defined for all vendors",
            "Contract templates with security clauses, executed vendor contracts"
        ),
        (
            "V-05",
            "Data protection requirements specified in vendor contracts (DPAs, data handling)",
            "Contract sections on data handling, Data Processing Agreements (DPAs), privacy clauses"
        ),
        (
            "V-06",
            "Incident response coordination requirements included in vendor contracts",
            "Incident notification SLAs, contact procedures, escalation paths in contracts"
        ),
        (
            "V-07",
            "Vendor access controls documented and enforced (least privilege)",
            "Access logs, permission matrices, SSO configurations, access request records"
        ),
        (
            "V-08",
            "Vendor accounts follow least privilege principle (minimal necessary permissions)",
            "Access reviews, role definitions, permission audits, access justifications"
        ),
        (
            "V-09",
            "Vendor access reviewed regularly (quarterly for high-risk vendors)",
            "Access review reports, approval records, review meeting notes, access certification logs"
        ),
        (
            "V-10",
            "High-risk vendors undergo annual comprehensive security reviews",
            "Review meeting notes, security assessment reports, vendor scorecards, review schedules"
        ),
        (
            "V-11",
            "Vendor security incidents tracked and reviewed (lessons learned)",
            "Incident logs, post-mortem reports, lessons learned documents, remediation plans"
        ),
        (
            "V-12",
            "Vendor security posture monitored (security ratings, breach notifications)",
            "Security rating subscriptions (BitSight, SecurityScorecard), monitoring reports, breach alerts"
        ),
        (
            "V-13",
            "Vendor SOC 2 / ISO 27001 compliance verified where applicable",
            "SOC 2 Type II reports, ISO 27001 certificates, attestation letters, compliance verification"
        ),
        (
            "V-14",
            "Vendor data retention and deletion requirements specified in contracts",
            "Contract data handling clauses, data deletion procedures, retention schedules"
        ),
        (
            "V-15",
            "Vendor security questionnaire responses validated (not blindly accepted)",
            "Validation notes, follow-up correspondence, independent verification evidence"
        ),
        (
            "V-16",
            "Vendor offboarding process includes security checklist (access revocation, data return)",
            "Offboarding checklists, access revocation logs, data deletion certificates"
        ),
        (
            "V-17",
            "Critical vendors have business continuity plans reviewed and tested",
            "BCP documents, disaster recovery plans, BCP test results, failover procedures"
        ),
        (
            "V-18",
            "Vendor security metrics tracked (% compliant vendors, review completion rates)",
            "Vendor compliance dashboard, metrics reports, KPI tracking, trend analysis"
        ),
    ]
    
    return requirements


def get_oss_management_requirements():
    """
    Domain 2: Open Source Software Management (18 requirements)
    
    Focus: OSS approval, inventory (SBOM), license tracking, abandoned dependencies
    
    Returns:
        list: List of tuples (ID, Requirement, Evidence Examples)
    """
    requirements = [
        (
            "O-01",
            "OSS usage policy exists and is approved (defines acceptable OSS practices)",
            "OSS policy document, approval signatures, policy distribution records"
        ),
        (
            "O-02",
            "OSS approval workflow defined (request → review → approve process)",
            "Workflow documentation, JIRA/ServiceNow workflow configs, process diagrams"
        ),
        (
            "O-03",
            "OSS approval process enforced before production use (not bypassed)",
            "Approval tickets, process audit logs, enforcement metrics, compliance reports"
        ),
        (
            "O-04",
            "OSS inventory maintained (SBOM) for all applications and services",
            "SBOM exports (CycloneDX, SPDX format), inventory databases, dependency manifests"
        ),
        (
            "O-05",
            "SBOM generated automatically in CI/CD pipeline (not manual)",
            "CI/CD pipeline configs, build logs showing SBOM generation, automation scripts"
        ),
        (
            "O-06",
            "OSS components tracked with specific version numbers (not version ranges)",
            "Dependency manifests (package.json, requirements.txt, pom.xml), lockfiles"
        ),
        (
            "O-07",
            "Transitive dependencies identified and tracked (dependencies of dependencies)",
            "Dependency tree reports, SCA tool outputs showing full dependency graph"
        ),
        (
            "O-08",
            "OSS license information captured in inventory for all dependencies",
            "License scan reports, SBOM license fields, license identification logs"
        ),
        (
            "O-09",
            "Abandoned/unmaintained OSS detected and tracked (staleness monitoring)",
            "Dependency staleness reports, EOL (end-of-life) tracking, maintainer activity checks"
        ),
        (
            "O-10",
            "OSS update policy defines acceptable staleness thresholds (e.g., <6 months since last update)",
            "Policy document with staleness thresholds, update frequency requirements"
        ),
        (
            "O-11",
            "OSS components reviewed for security before approval (not rubber-stamped)",
            "Security review checklists, approval criteria, security assessment notes"
        ),
        (
            "O-12",
            "Restricted licenses flagged during approval process (GPL, AGPL, SSPL)",
            "License allowlist/blocklist, flagging rules in approval workflow, legal review triggers"
        ),
        (
            "O-13",
            "OSS alternatives evaluated before approval (build vs. buy vs. OSS decision)",
            "Evaluation notes, decision records, alternative analysis documents"
        ),
        (
            "O-14",
            "OSS contribution guidelines exist if organization contributes to OSS projects",
            "Contribution policy, CLA (Contributor License Agreement) templates, approval process"
        ),
        (
            "O-15",
            "OSS forks minimized and justified (upstream contributions preferred)",
            "Fork justifications, upstream contribution plans, fork maintenance records"
        ),
        (
            "O-16",
            "OSS components with known high-risk maintainers flagged for review",
            "Maintainer reputation checks, risk assessments, red flag criteria"
        ),
        (
            "O-17",
            "OSS inventory reviewed regularly (quarterly minimum) for accuracy",
            "Inventory review reports, approval records, accuracy audit logs"
        ),
        (
            "O-18",
            "OSS metrics tracked (% approved components, inventory coverage, SBOM completeness)",
            "OSS compliance dashboard, metrics reports, KPI tracking"
        ),
    ]
    
    return requirements

def get_dependency_vulnerability_requirements():
    """
    Domain 3: Dependency Vulnerability Management (18 requirements)
    
    Focus: SCA tools, vulnerability detection, remediation SLAs, lockfiles
    
    Returns:
        list: List of tuples (ID, Requirement, Evidence Examples)
    """
    requirements = [
        (
            "D-01",
            "SCA (Software Composition Analysis) tool deployed and scanning dependencies",
            "SCA tool configurations (Snyk, Dependabot, WhiteSource, Sonatype), deployment records"
        ),
        (
            "D-02",
            "SCA scans run automatically on every build/commit (not manual)",
            "CI/CD pipeline configs, scan logs, automation evidence, build failure logs"
        ),
        (
            "D-03",
            "SCA tool integrated with developer workflow (IDE plugins, PR checks)",
            "IDE plugin configurations, PR check results, developer tool screenshots"
        ),
        (
            "D-04",
            "Vulnerability alerts delivered to responsible teams (not ignored)",
            "Alert routing configurations, notification logs, Slack/email integrations"
        ),
        (
            "D-05",
            "Vulnerability remediation SLAs defined by severity (Critical/High/Medium/Low)",
            "SLA policy document, severity definitions, remediation timelines"
        ),
        (
            "D-06",
            "Critical vulnerabilities have <7 day remediation SLA (e.g., Log4Shell scenarios)",
            "SLA policy with critical timelines, compliance tracking, escalation procedures"
        ),
        (
            "D-07",
            "High vulnerabilities have <30 day remediation SLA",
            "SLA policy with high severity timelines, tracking metrics, compliance reports"
        ),
        (
            "D-08",
            "Vulnerability triage process documented (risk assessment, prioritization)",
            "Triage workflow documentation, decision criteria, risk scoring methodology"
        ),
        (
            "D-09",
            "False positives documented with suppression justification (not silently ignored)",
            "Suppression records, justification notes, approval logs, suppression review process"
        ),
        (
            "D-10",
            "Dependency update process documented (testing requirements, rollback plans)",
            "Update procedures, testing requirements, rollback documentation, change management"
        ),
        (
            "D-11",
            "Dependency updates tested before production deployment (not YOLO updates)",
            "Test results, staging environment logs, regression test evidence, QA sign-offs"
        ),
        (
            "D-12",
            "Transitive dependency vulnerabilities addressed (not just direct dependencies)",
            "Transitive vulnerability remediation records, dependency tree analysis, patch logs"
        ),
        (
            "D-13",
            "Dependency pinning used (lockfiles required: package-lock.json, Pipfile.lock, etc.)",
            "Lockfile examples, policy enforcement, CI/CD checks for lockfile presence"
        ),
        (
            "D-14",
            "Dependency version ranges minimized (avoid wildcards like '^1.0.0' or '~1.0.0')",
            "Dependency manifests showing pinned versions, policy on version ranges"
        ),
        (
            "D-15",
            "Emergency patching process for critical vulnerabilities (Log4Shell-style response plan)",
            "Emergency response plans, drill records, war room procedures, communication plans"
        ),
        (
            "D-16",
            "Vulnerability remediation tracked with metrics (MTTR, remediation rates)",
            "Remediation dashboards, MTTR (Mean Time To Remediate) metrics, trend reports"
        ),
        (
            "D-17",
            "Unpatched vulnerabilities require formal risk acceptance (no silent ignoring)",
            "Risk acceptance forms, approval signatures, risk register, acceptance expiry tracking"
        ),
        (
            "D-18",
            "SCA tool effectiveness reviewed regularly (false negative testing, coverage checks)",
            "Tool effectiveness reports, benchmark tests, false negative rate analysis"
        ),
    ]
    
    return requirements


def get_third_party_code_review_requirements():
    """
    Domain 4: Third-Party Code Review & Integration (18 requirements)
    
    Focus: Integration security, API assessment, package integrity, supply chain attacks
    
    Returns:
        list: List of tuples (ID, Requirement, Evidence Examples)
    """
    requirements = [
        (
            "T-01",
            "Third-party code integration policy exists (defines security review requirements)",
            "Integration policy document, approval records, security review requirements"
        ),
        (
            "T-02",
            "Security review required for all third-party integrations before production",
            "Review checklists, approval workflow, integration assessment records"
        ),
        (
            "T-03",
            "Third-party APIs assessed for security risks (authentication, authorization, data exposure)",
            "API security assessments, risk ratings, vulnerability scan results"
        ),
        (
            "T-04",
            "API authentication mechanisms reviewed (OAuth, API keys, mutual TLS)",
            "Authentication config reviews, security assessment notes, credential management"
        ),
        (
            "T-05",
            "API authorization controls verified (least privilege, scope limitations)",
            "Authorization testing results, access reviews, scope definitions, permission audits"
        ),
        (
            "T-06",
            "Third-party API rate limiting and abuse prevention assessed",
            "Rate limiting configurations, abuse detection logs, throttling evidence"
        ),
        (
            "T-07",
            "SDK/library source code reviewed where feasible (especially for critical dependencies)",
            "Code review reports, static analysis results, security findings"
        ),
        (
            "T-08",
            "Third-party container images scanned for vulnerabilities (Trivy, Clair, etc.)",
            "Container scan reports, vulnerability findings, remediation evidence"
        ),
        (
            "T-09",
            "Container base images from trusted sources only (official registries, verified publishers)",
            "Image provenance records, registry configurations, trusted source lists"
        ),
        (
            "T-10",
            "Package integrity verified before installation (checksums, GPG signatures)",
            "Package verification configurations (GPG setup, SHA256 checks), verification logs"
        ),
        (
            "T-11",
            "Package manager security features enabled (npm audit, pip-audit, bundler-audit)",
            "Package manager configurations, audit command integration, CI/CD security checks"
        ),
        (
            "T-12",
            "Typosquatting protection implemented (package name validation, allowlists)",
            "Package allowlists, validation rules, namespace protection, monitoring alerts"
        ),
        (
            "T-13",
            "Third-party code isolated from critical systems (sandboxing, containerization)",
            "Architecture diagrams, isolation configurations, namespace separation, security boundaries"
        ),
        (
            "T-14",
            "Vendor-provided code deployed with least privilege (minimal permissions)",
            "Permission configurations, role definitions, privilege audits, access reviews"
        ),
        (
            "T-15",
            "Code escrow agreements in place for critical vendors (optional, for business continuity)",
            "Escrow agreements, escrow deposit verification, code access procedures"
        ),
        (
            "T-16",
            "Third-party integration monitoring in place (API calls, error rates, anomalies)",
            "Integration logs, monitoring dashboards, alert configurations, anomaly detection"
        ),
        (
            "T-17",
            "Supply chain attack mitigations documented (dependency confusion, typosquatting, compromised packages)",
            "Mitigation strategies, security controls, attack surface analysis, prevention measures"
        ),
        (
            "T-18",
            "Third-party integration security reviewed regularly (annual reviews minimum)",
            "Integration review reports, approval records, security reassessments, audit logs"
        ),
    ]
    
    return requirements


def get_license_compliance_requirements():
    """
    Domain 5: License Compliance & Legal Risk (18 requirements)
    
    Focus: License identification, GPL/copyleft management, legal compliance
    
    Returns:
        list: List of tuples (ID, Requirement, Evidence Examples)
    """
    requirements = [
        (
            "L-01",
            "License compliance policy exists and is approved by legal counsel",
            "Policy document, legal approval signatures, policy distribution records"
        ),
        (
            "L-02",
            "License scanning tool deployed (FOSSA, Black Duck, WhiteSource, etc.)",
            "License scan tool configurations, deployment records, integration evidence"
        ),
        (
            "L-03",
            "License scans run automatically on every build (not manual/ad-hoc)",
            "CI/CD configurations, scan logs, build pipeline evidence, automation scripts"
        ),
        (
            "L-04",
            "All dependency licenses identified and tracked (no unknown licenses)",
            "License scan reports, SBOM license data, license inventory, unknown license alerts"
        ),
        (
            "L-05",
            "License allowlist and blocklist defined (approved vs. restricted licenses)",
            "Allowlist/blocklist documents, approval criteria, legal guidance, license policies"
        ),
        (
            "L-06",
            "Copyleft licenses flagged for legal review (GPL, AGPL, MPL, EPL)",
            "Flagging rules in scan tools, legal review records, copyleft license alerts"
        ),
        (
            "L-07",
            "License compatibility assessed (GPL + proprietary conflicts detected)",
            "Compatibility matrix, conflict reports, license interaction analysis, legal opinions"
        ),
        (
            "L-08",
            "Viral license contamination prevented (GPL/AGPL code isolated from proprietary code)",
            "Isolation strategies, architecture reviews, code segregation, legal firewalls"
        ),
        (
            "L-09",
            "Commercial licenses tracked with usage limits (seat counts, deployment restrictions)",
            "License tracking system, usage reports, compliance monitoring, license renewals"
        ),
        (
            "L-10",
            "License violation detection automated (blocklist enforcement, conflict detection)",
            "Violation detection rules, alert configurations, automated blocking, CI/CD gates"
        ),
        (
            "L-11",
            "License violations remediated with documented plan (not ignored)",
            "Remediation plans, resolution records, alternative selection, legal consultation"
        ),
        (
            "L-12",
            "Legal review required for restrictive licenses (AGPL, SSPL, Business Source License)",
            "Legal review requests, approval records, risk assessments, legal opinions"
        ),
        (
            "L-13",
            "Open source attribution maintained (LICENSE, NOTICE, THIRD_PARTY_LICENSES files)",
            "Attribution files in repositories, automated generation scripts, distribution packages"
        ),
        (
            "L-14",
            "Third-party license notices included in product distributions (legal requirement)",
            "Distribution packages, license bundling, notice files, compliance documentation"
        ),
        (
            "L-15",
            "License compliance verified before production release (release gate)",
            "Release checklists, compliance sign-offs, license approval records, gate evidence"
        ),
        (
            "L-16",
            "License audit trails maintained (license changes, approvals, violations)",
            "Audit logs, compliance history, license change tracking, approval records"
        ),
        (
            "L-17",
            "License compliance training provided to developers (awareness of GPL, AGPL, etc.)",
            "Training records, completion certificates, training materials, quiz results"
        ),
        (
            "L-18",
            "License compliance metrics tracked (% compliant, violations, legal reviews)",
            "Compliance dashboards, metrics reports, KPI tracking, trend analysis"
        ),
    ]
    
    return requirements

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS
# ============================================================================

def create_instructions_sheet(wb, styles):
    """
    Create the Instructions sheet with assessment guidance and methodology.
    
    This sheet provides:
    - Assessment overview and philosophy (Feynman principles)
    - Domain-specific instructions
    - Evidence requirements
    - Anti-cargo-cult reminders
    - Scoring guidance
    
    Args:
        wb: Workbook object
        styles: Style definitions dictionary
    """
    ws = wb.create_sheet("Instructions", 0)
    
    # Set column widths for instructions sheet
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 90
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "ISMS Control A.8.28.4 - Third-Party & OSS Assessment Instructions"
    cell.font = styles['header_font']
    cell.fill = styles['header_fill']
    cell.alignment = styles['align_center']
    current_row += 2
    
    # Assessment Overview
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "1. ASSESSMENT OVERVIEW"
    cell.font = styles['subheader_font']
    cell.fill = styles['subheader_fill']
    cell.alignment = styles['align_left']
    current_row += 1
    
    overview_text = """This assessment evaluates your organization's management of third-party dependencies, open source software (OSS), vendor-provided components, and external code integrations.

Supply chain security is critical - major breaches like SolarWinds, Log4Shell, and Codecov exploited weaknesses in third-party software management. This assessment ensures you have proper controls to prevent similar incidents.

PHILOSOPHY (Richard Feynman): "The first principle is that you must not fool yourself—and you are the easiest person to fool."

This assessment focuses on ACTUAL PRACTICES, not just documented policies. Having a vendor questionnaire ≠ vendors are secure. Having an SCA tool ≠ vulnerabilities are fixed. We verify what actually works, not what's written down.

ANTI-CARGO-CULT REMINDER: Don't just go through the motions. If you mark something "Implemented," you should be able to demonstrate it working right now. If you can't, it's "Partially Implemented" or "Not Implemented." Honesty in assessment leads to real security improvements."""
    
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = overview_text
    cell.font = styles['normal_font']
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[current_row].height = 200
    current_row += 2
    
    # Assessment Domains
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "2. ASSESSMENT DOMAINS (90 Requirements Total)"
    cell.font = styles['subheader_font']
    cell.fill = styles['subheader_fill']
    cell.alignment = styles['align_left']
    current_row += 1
    
    domains_text = """This assessment covers 5 domains with 18 requirements each:

DOMAIN 1: VENDOR SECURITY ASSESSMENT (Sheet: Vendor_Security_Assessment)
- Focus: Vendor due diligence, contracts, access controls, risk management
- Key Question: Do we know which vendors have access to our systems/data and are they secure?
- Common Pitfall: Vendor onboarding without security review (blind trust)

DOMAIN 2: OPEN SOURCE SOFTWARE MANAGEMENT (Sheet: OSS_Management)
- Focus: OSS approval, inventory (SBOM), license tracking, abandoned dependencies
- Key Question: Do we maintain an accurate inventory of all open source dependencies?
- Common Pitfall: "npm install" without review, no SBOM, stale dependencies ignored

DOMAIN 3: DEPENDENCY VULNERABILITY MANAGEMENT (Sheet: Dependency_Vulnerability_Mgmt)
- Focus: SCA tools, vulnerability detection, remediation SLAs, lockfiles
- Key Question: Are dependencies scanned for vulnerabilities and patched within SLAs?
- Common Pitfall: Ignoring SCA tool alerts, no SLA for patching (Log4Shell scenarios)

DOMAIN 4: THIRD-PARTY CODE REVIEW & INTEGRATION (Sheet: Third_Party_Code_Review)
- Focus: Integration security, API assessment, package integrity, supply chain attacks
- Key Question: Are third-party integrations security reviewed before production?
- Common Pitfall: Vendor code deployed without review, no package verification (typosquatting risk)

DOMAIN 5: LICENSE COMPLIANCE & LEGAL RISK (Sheet: License_Compliance)
- Focus: License identification, GPL/copyleft management, legal compliance
- Key Question: Are all dependency licenses identified and compatible with our code?
- Common Pitfall: GPL contamination (mixing GPL code with proprietary), missing attribution files"""
    
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = domains_text
    cell.font = styles['normal_font']
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[current_row].height = 280
    current_row += 2
    
    # How to Complete
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "3. HOW TO COMPLETE THIS ASSESSMENT"
    cell.font = styles['subheader_font']
    cell.fill = styles['subheader_fill']
    cell.alignment = styles['align_left']
    current_row += 1
    
    completion_text = """STEP 1: Review Each Domain Sheet
- Navigate to each domain sheet (Vendor_Security_Assessment, OSS_Management, etc.)
- Read each requirement carefully
- Consider the evidence examples provided

STEP 2: Assess Implementation Status
For each requirement, select the appropriate status from the dropdown in Column C:
- "Implemented" = Requirement fully met, evidence available, working as intended
- "Partially Implemented" = Requirement partially met, some gaps exist
- "Not Implemented" = Requirement not met, significant gaps
- "N/A" = Requirement not applicable to your environment (use sparingly)

STEP 3: Document Evidence (Column D)
Provide specific evidence references:
- Good: "Vendor security policy v2.1 approved 15.03.2024, SharePoint link: [URL]"
- Good: "Snyk SCA scans in CI/CD, see Jenkins pipeline config line 45"
- Bad: "We have this"
- Bad: "Compliant"

STEP 4: Add Comments/Context (Column E)
Use this for:
- Clarifying partial implementations
- Noting planned improvements
- Explaining N/A selections
- Documenting compensating controls

STEP 5: Track Evidence (Evidence_Register Sheet)
- Log all evidence in the centralized Evidence Register
- Include evidence location, owner, verification status
- This creates an audit trail

STEP 6: Identify Gaps (Gap_Analysis Sheet)
- For "Not Implemented" or "Partially Implemented" items, create a gap entry
- Assign owner, target date, remediation plan
- Track gap closure

STEP 7: Review Summary (Summary_Dashboard Sheet)
- Check overall compliance percentage
- Review domain-specific compliance
- Identify highest-risk gaps

STEP 8: Obtain Approval (Approval_Sign_Off Sheet)
- Submit completed assessment for review
- Obtain signatures from required approvers
- Archive completed workbook"""
    
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = completion_text
    cell.font = styles['normal_font']
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[current_row].height = 380
    current_row += 2
    
    # Evidence Requirements
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "4. EVIDENCE REQUIREMENTS & EXAMPLES"
    cell.font = styles['subheader_font']
    cell.fill = styles['subheader_fill']
    cell.alignment = styles['align_left']
    current_row += 1
    
    evidence_text = """WHAT MAKES GOOD EVIDENCE?

GOOD Evidence Characteristics:
✓ Specific: References exact documents, systems, or configurations
✓ Verifiable: Can be independently validated by an auditor
✓ Current: Recent evidence (within last 12 months for most items)
✓ Traceable: Clear location/link/reference provided

EVIDENCE TYPE EXAMPLES:

Policy Documents: "Vendor Security Assessment Policy v1.2, approved 10.01.2025, stored at: SharePoint/Policies/Vendor_Security.pdf"

Tool Configurations: "Snyk SCA integration, see GitHub Actions workflow: .github/workflows/security-scan.yml, lines 25-40"

Scan Reports: "Dependency scan report dated 05.01.2025, 147 dependencies scanned, 3 high vulnerabilities found, see Jenkins build #1234"

Process Evidence: "OSS approval workflow in JIRA, project key: SEC, approval workflow screenshot dated 03.01.2025"

Contracts: "Vendor contract with Acme Corp, executed 12.11.2024, security clauses Section 8.1-8.5, contract ID: VC-2024-045"

Logs/Records: "Vendor access review completed 20.12.2024, 12 vendors reviewed, access log export attached"

Metrics/Dashboards: "SCA vulnerability remediation dashboard, MTTR = 4.2 days for critical vulns, dashboard URL: [internal]"

BAD Evidence (Avoid These):
✗ "We do this" - Not specific, not verifiable
✗ "Compliant" - Meaningless without details
✗ "Yes" - Provides no evidence reference
✗ "See attached" - Without clear file reference
✗ "It's in the system" - Which system? Where?"""
    
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = evidence_text
    cell.font = styles['normal_font']
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[current_row].height = 320
    current_row += 2
    
    # Anti-Cargo-Cult Section
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "5. ANTI-CARGO-CULT: COMMON FALSE POSITIVES"
    cell.font = styles['subheader_font']
    cell.fill = styles['subheader_fill']
    cell.alignment = styles['align_left']
    current_row += 1
    
    cargo_cult_text = """These look like "Implemented" but often aren't (be honest!):

\u274C "We have a vendor questionnaire" ≠ Vendors are secure
Reality Check: Do you validate questionnaire answers? Are they tailored to risk level? Do you follow up on concerning responses? If you just file them away, mark "Partially Implemented."

\u274C "We scan for vulnerabilities" ≠ Vulnerabilities are fixed
Reality Check: What's your mean time to remediate critical vulns? Are SLAs actually met? Or do you have 500 unaddressed findings? Scanning without remediation is theater.

\u274C "We track licenses" ≠ License compliance
Reality Check: Do you remediate violations? Is legal involved in GPL decisions? Are attribution files actually maintained? Or is it just a report nobody reads?

\u274C "We have an OSS policy" ≠ OSS is managed
Reality Check: Is the approval process enforced? Do teams bypass it? Is your SBOM accurate and complete? Or is the policy a dusty PDF?

\u274C "We use lockfiles" ≠ Dependencies are pinned
Reality Check: Are lockfiles checked into version control? Are they generated automatically? Or do developers delete them when they cause conflicts?

\u274C "We review vendor contracts" ≠ Contracts have security requirements
Reality Check: Do contracts actually include security clauses? Or is it just boilerplate with no enforcement? Can you show me the security sections?

\u274C "Developers know about security" ≠ Security is practiced
Reality Check: Do developers actually follow secure coding practices? Or did they attend one training and forget it? Evidence speaks louder than assumptions.

FEYNMAN'S CHALLENGE: "What I cannot create, I do not understand."
If you can't demonstrate the control working RIGHT NOW, you don't have the control. Be honest with yourself."""
    
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = cargo_cult_text
    cell.font = styles['normal_font']
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[current_row].height = 340
    current_row += 2
    
    # Scoring Guidance
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "6. SCORING GUIDANCE & SUCCESS CRITERIA"
    cell.font = styles['subheader_font']
    cell.fill = styles['subheader_fill']
    cell.alignment = styles['align_left']
    current_row += 1
    
    scoring_text = """COMPLIANCE CALCULATION:
Compliance % = (Implemented + Partially Implemented × 0.5) / (Total Requirements - N/A) × 100

Example: Domain has 18 requirements
- 10 Implemented = 10 points
- 4 Partially Implemented = 2 points (4 × 0.5)
- 3 Not Implemented = 0 points
- 1 N/A = excluded from calculation
- Compliance % = (10 + 2) / (18 - 1) × 100 = 70.6%

TARGET COMPLIANCE LEVELS:
- Overall Assessment: ≥ 70% (minimum acceptable)
- Vendor Security Assessment: ≥ 65% (contracts take time)
- OSS Management: ≥ 75% (foundational requirement)
- Dependency Vulnerability Mgmt: ≥ 80% (critical for security)
- Third-Party Code Review: ≥ 65% (depends on complexity)
- License Compliance: ≥ 70% (legal requirement)

MATURITY ROADMAP:
Year 1 (Initial): 70-75% overall compliance - Basic controls in place
Year 2 (Optimization): 80-85% overall compliance - Controls effective and measurable
Year 3+ (Mature): 90%+ overall compliance - Continuous improvement culture

CRITICAL REQUIREMENTS (Must be "Implemented"):
These cannot be "Partially Implemented" or "Not Implemented" regardless of overall score:
- O-04: OSS inventory maintained (SBOM) - You must know what you're running
- D-02: SCA scans run automatically - Manual scans don't scale
- D-06: Critical vulnerabilities <7 day SLA - Log4Shell cannot wait 30 days
- L-04: All dependency licenses identified - Ignorance is not a legal defense
- L-07: License compatibility assessed - GPL contamination is expensive to fix

RED FLAGS (Immediate CISO Escalation):
- No OSS inventory (SBOM) exists - You're flying blind
- No SCA tool deployed - Vulnerabilities are invisible
- Critical vulnerabilities >30 days unpatched - Active exploitation risk
- GPL contamination discovered (GPL + proprietary code) - Legal liability
- Vendor breach involving company data - Incident response required"""
    
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = scoring_text
    cell.font = styles['normal_font']
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[current_row].height = 380
    current_row += 2
    
    # Resources
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "7. RESOURCES & CONTACTS"
    cell.font = styles['subheader_font']
    cell.fill = styles['subheader_fill']
    cell.alignment = styles['align_left']
    current_row += 1
    
    resources_text = """ASSESSMENT SUPPORT:
- Application Security Lead: [Contact details]
- CISO Office: [Contact details]
- Legal Counsel (License Compliance): [Contact details]
- Procurement (Vendor Contracts): [Contact details]

RELATED POLICIES & DOCUMENTS:
- ISMS-POL-A.8.28 - Secure Coding Policy (Master Policy)
- ISMS-POL-A.8.28-S2.4 - Third-Party & OSS Management Policy
- ISMS-IMP-A.8.28.1 - SDLC Assessment
- ISMS-IMP-A.8.28.2 - Standards & Tools Assessment

INDUSTRY FRAMEWORKS:
- SLSA (Supply-chain Levels for Software Artifacts): https://slsa.dev
- NIST SSDF (Secure Software Development Framework): NIST SP 800-218
- OWASP Dependency-Check: https://owasp.org/www-project-dependency-check/
- CycloneDX SBOM Standard: https://cyclonedx.org

QUESTIONS OR ISSUES?
Contact the Application Security team or use the assessment feedback mechanism."""
    
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = resources_text
    cell.font = styles['normal_font']
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[current_row].height = 200
    current_row += 2
    
    # Footer
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "END OF INSTRUCTIONS - Proceed to Domain Assessment Sheets"
    cell.font = Font(name='Calibri', size=10, bold=True, italic=True)
    cell.alignment = styles['align_center']
    
    print("✓ Instructions sheet created")


def create_domain_assessment_sheet(wb, sheet_name, domain_title, requirements, styles, validations, column_widths):
    """
    Create a domain assessment sheet with requirements and input fields.
    
    Args:
        wb: Workbook object
        sheet_name: Name of the sheet (e.g., "Vendor_Security_Assessment")
        domain_title: Display title for the domain
        requirements: List of tuples (ID, Requirement, Evidence Examples)
        styles: Style definitions dictionary
        validations: Data validation objects dictionary
        column_widths: Column width definitions
    """
    ws = wb.create_sheet(sheet_name)
    
    # Set column widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = domain_title
    cell.font = styles['header_font']
    cell.fill = styles['header_fill']
    cell.alignment = styles['align_center']
    cell.border = styles['thick_border']
    current_row += 1
    
    # Metadata
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = "Assessment Date:"
    ws[f'A{current_row}'].font = styles['bold_font']
    ws.merge_cells(f'C{current_row}:D{current_row}')
    ws[f'C{current_row}'] = datetime.now().strftime('%d.%m.%Y')
    ws[f'C{current_row}'].fill = styles['input_fill']
    
    ws[f'E{current_row}'] = "Assessor:"
    ws[f'E{current_row}'].font = styles['bold_font']
    ws[f'F{current_row}'].fill = styles['input_fill']
    current_row += 1
    
    # Column Headers
    headers = ["ID", "Requirement", "Implementation Status", "Evidence Reference", "Comments", "✓"]
    set_header_row(ws, current_row, headers, styles)
    current_row += 1
    
    # Freeze header rows
    freeze_header_rows(ws, current_row)
    
    # Data rows
    start_data_row = current_row
    for req_id, requirement, evidence_examples in requirements:
        # Column A: ID
        ws[f'A{current_row}'] = req_id
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='top')
        
        # Column B: Requirement (with evidence examples in smaller text)
        full_text = f"{requirement}\n\nEvidence Examples: {evidence_examples}"
        ws[f'B{current_row}'] = full_text
        ws[f'B{current_row}'].alignment = styles['align_left']
        ws.row_dimensions[current_row].height = 60
        
        # Column C: Implementation Status (dropdown)
        ws[f'C{current_row}'].fill = styles['input_fill']
        
        # Column D: Evidence Reference (input)
        ws[f'D{current_row}'].fill = styles['input_fill']
        
        # Column E: Comments (input)
        ws[f'E{current_row}'].fill = styles['input_fill']
        
        # Column F: Compliance indicator (formula)
        formula = f'=IF(C{current_row}="Implemented","✓",IF(C{current_row}="Partially Implemented","◐",IF(C{current_row}="Not Implemented","✗",IF(C{current_row}="N/A","—",""))))'
        ws[f'F{current_row}'] = formula
        ws[f'F{current_row}'].alignment = styles['align_center']
        
        current_row += 1
    
    end_data_row = current_row - 1
    
    # Apply formatting and validation
    apply_standard_formatting(ws, start_data_row, end_data_row, styles, column_widths)
    apply_validation_to_range(ws, validations['implementation_status'], start_data_row, end_data_row, 'C')
    
    # Conditional formatting for Column F (compliance indicator)
    from openpyxl.formatting.rule import CellIsRule
    
    ws.conditional_formatting.add(
        f'F{start_data_row}:F{end_data_row}',
        CellIsRule(operator='equal', formula=['"✓"'], fill=styles['status_green'])
    )
    ws.conditional_formatting.add(
        f'F{start_data_row}:F{end_data_row}',
        CellIsRule(operator='equal', formula=['"◐"'], fill=styles['status_yellow'])
    )
    ws.conditional_formatting.add(
        f'F{start_data_row}:F{end_data_row}',
        CellIsRule(operator='equal', formula=['"✗"'], fill=styles['status_red'])
    )
    ws.conditional_formatting.add(
        f'F{start_data_row}:F{end_data_row}',
        CellIsRule(operator='equal', formula=['"—"'], fill=styles['status_gray'])
    )
    
    print(f"✓ {sheet_name} sheet created ({len(requirements)} requirements)")

# ============================================================================
# SECTION 5: SUPPORTING SHEETS (Summary, Evidence, Gap, Approval)
# ============================================================================

def create_summary_dashboard(wb, styles):
    """Create Summary Dashboard with aggregated metrics."""
    ws = wb.create_sheet("Summary_Dashboard")
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "Assessment Summary Dashboard"
    cell.font = styles['header_font']
    cell.fill = styles['header_fill']
    cell.alignment = styles['align_center']
    current_row += 2
    
    # Assessment Info
    ws[f'A{current_row}'] = "Assessment Date:"
    ws[f'A{current_row}'].font = styles['bold_font']
    ws[f'B{current_row}'] = datetime.now().strftime('%d.%m.%Y')
    ws[f'D{current_row}'] = "Assessor:"
    ws[f'D{current_row}'].font = styles['bold_font']
    ws[f'E{current_row}'].fill = styles['input_fill']
    current_row += 2
    
    # Domain Compliance Table
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "Domain Compliance Overview"
    cell.font = styles['subheader_font']
    cell.fill = styles['subheader_fill']
    cell.alignment = styles['align_left']
    current_row += 1
    
    # Headers
    headers = ["Domain", "Total", "Impl.", "Partial", "Not Impl.", "N/A", "Compliance %", "Status"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=idx)
        cell.value = header
        cell.font = styles['column_header_font']
        cell.fill = styles['column_header_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['thin_border']
    current_row += 1
    
    # Domain rows
    domains = [
        ("Vendor Security Assessment", "Vendor_Security_Assessment", 18),
        ("OSS Management", "OSS_Management", 18),
        ("Dependency Vulnerability Mgmt", "Dependency_Vulnerability_Mgmt", 18),
        ("Third-Party Code Review", "Third_Party_Code_Review", 18),
        ("License Compliance", "License_Compliance", 18),
    ]
    
    domain_start_row = current_row
    for domain_name, sheet_name, total_reqs in domains:
        ws[f'A{current_row}'] = domain_name
        ws[f'B{current_row}'] = total_reqs
        
        # Count formulas
        ws[f'C{current_row}'] = f'=COUNTIF({sheet_name}!C4:C21,"Implemented")'
        ws[f'D{current_row}'] = f'=COUNTIF({sheet_name}!C4:C21,"Partially Implemented")'
        ws[f'E{current_row}'] = f'=COUNTIF({sheet_name}!C4:C21,"Not Implemented")'
        ws[f'F{current_row}'] = f'=COUNTIF({sheet_name}!C4:C21,"N/A")'
        
        # Compliance % formula
        ws[f'G{current_row}'] = f'=IF((B{current_row}-F{current_row})=0,0,(C{current_row}+D{current_row}*0.5)/(B{current_row}-F{current_row}))'
        ws[f'G{current_row}'].number_format = '0.0%'
        
        # Status formula (traffic light)
        ws[f'H{current_row}'] = f'=IF(G{current_row}>=0.8,"🟢 Green",IF(G{current_row}>=0.5,"🟡 Yellow","🔴 Red"))'
        
        # Apply borders
        for col in range(1, 9):
            ws.cell(row=current_row, column=col).border = styles['thin_border']
            ws.cell(row=current_row, column=col).alignment = styles['align_center']
        
        current_row += 1
    
    domain_end_row = current_row - 1
    
    # Overall totals
    ws[f'A{current_row}'] = "OVERALL TOTAL"
    ws[f'A{current_row}'].font = styles['bold_font']
    ws[f'B{current_row}'] = f'=SUM(B{domain_start_row}:B{domain_end_row})'
    ws[f'C{current_row}'] = f'=SUM(C{domain_start_row}:C{domain_end_row})'
    ws[f'D{current_row}'] = f'=SUM(D{domain_start_row}:D{domain_end_row})'
    ws[f'E{current_row}'] = f'=SUM(E{domain_start_row}:E{domain_end_row})'
    ws[f'F{current_row}'] = f'=SUM(F{domain_start_row}:F{domain_end_row})'
    ws[f'G{current_row}'] = f'=IF((B{current_row}-F{current_row})=0,0,(C{current_row}+D{current_row}*0.5)/(B{current_row}-F{current_row}))'
    ws[f'G{current_row}'].number_format = '0.0%'
    ws[f'H{current_row}'] = f'=IF(G{current_row}>=0.7,"🟢 Pass","🔴 Fail")'
    
    for col in range(1, 9):
        ws.cell(row=current_row, column=col).font = styles['bold_font']
        ws.cell(row=current_row, column=col).border = styles['thick_border']
        ws.cell(row=current_row, column=col).alignment = styles['align_center']
    
    current_row += 2
    
    # Key Statistics
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "Key Statistics"
    cell.font = styles['subheader_font']
    cell.fill = styles['subheader_fill']
    current_row += 1
    
    stats = [
        ("Total Requirements:", f'=B{domain_end_row + 1}'),
        ("Overall Compliance %:", f'=G{domain_end_row + 1}'),
        ("Critical Gaps (Not Impl.):", f'=E{domain_end_row + 1}'),
        ("Evidence Completeness:", '=COUNTA(Evidence_Register!D2:D200)/COUNTA(Evidence_Register!A2:A200)'),
    ]
    
    for label, formula in stats:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = styles['bold_font']
        ws[f'B{current_row}'] = formula
        if "%" in label:
            ws[f'B{current_row}'].number_format = '0.0%'
        current_row += 1
    
    print("✓ Summary_Dashboard sheet created")


def create_evidence_register(wb, styles, validations):
    """Create Evidence Register for tracking all evidence."""
    ws = wb.create_sheet("Evidence_Register")
    
    # Column widths
    widths = {'A': 10, 'B': 40, 'C': 15, 'D': 35, 'E': 12, 'F': 20, 'G': 15, 'H': 20, 'I': 12, 'J': 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    # Title
    current_row = add_title_section(ws, 1, "Evidence Register", styles)
    
    # Headers
    headers = ["Req ID", "Requirement Description", "Evidence Type", "Evidence Location", 
                "Evidence Date", "Evidence Owner", "Verification Status", "Verifier Name", 
                "Verification Date", "Notes"]
    set_header_row(ws, current_row, headers, styles)
    freeze_header_rows(ws, current_row + 1)
    
    # Add data validation for Evidence Type and Verification Status
    ws.add_data_validation(validations['evidence_type'])
    ws.add_data_validation(validations['verification_status'])
    
    for row in range(current_row + 1, current_row + 101):
        validations['evidence_type'].add(f'C{row}')
        validations['verification_status'].add(f'G{row}')
        
        # Apply input fill to editable columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws[f'{col}{row}'].fill = styles['input_fill']
            ws[f'{col}{row}'].border = styles['thin_border']
    
    print("✓ Evidence_Register sheet created")


def create_gap_analysis(wb, styles, validations):
    """Create Gap Analysis sheet for tracking remediation."""
    ws = wb.create_sheet("Gap_Analysis")
    
    # Column widths
    widths = {'A': 10, 'B': 25, 'C': 10, 'D': 35, 'E': 30, 'F': 30, 'G': 12, 'H': 35, 
              'I': 20, 'J': 12, 'K': 15, 'L': 12, 'M': 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    # Title
    current_row = add_title_section(ws, 1, "Gap Analysis & Remediation Tracking", styles)
    
    # Headers
    headers = ["Gap ID", "Domain", "Req ID", "Gap Description", "Current State", "Target State",
               "Risk Level", "Remediation Plan", "Owner", "Target Date", "Status", "Closure Date", "Notes"]
    set_header_row(ws, current_row, headers, styles)
    freeze_header_rows(ws, current_row + 1)
    
    # Add data validations
    ws.add_data_validation(validations['domain'])
    ws.add_data_validation(validations['priority'])
    ws.add_data_validation(validations['gap_status'])
    
    for row in range(current_row + 1, current_row + 101):
        # Gap ID auto-increment
        if row == current_row + 1:
            ws[f'A{row}'] = "GAP-001"
        else:
            ws[f'A{row}'] = f'="GAP-"&TEXT(ROW()-{current_row},"000")'
        
        # Apply validations
        validations['domain'].add(f'B{row}')
        validations['priority'].add(f'G{row}')
        validations['gap_status'].add(f'K{row}')
        
        # Apply input fill
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{row}'].fill = styles['input_fill']
            ws[f'{col}{row}'].border = styles['thin_border']
    
    # Conditional formatting for overdue items
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        f'J{current_row + 1}:J{current_row + 100}',
        CellIsRule(operator='lessThan', formula=[f'TODAY()'], fill=styles['status_red'])
    )
    
    print("✓ Gap_Analysis sheet created")


def create_approval_signoff(wb, styles, validations):
    """Create Approval Sign-Off sheet for formal approval."""
    ws = wb.create_sheet("Approval_Sign_Off")
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "Assessment Approval & Sign-Off"
    cell.font = styles['header_font']
    cell.fill = styles['header_fill']
    cell.alignment = styles['align_center']
    current_row += 2
    
    # Assessment Summary
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "Assessment Summary"
    cell.font = styles['subheader_font']
    cell.fill = styles['subheader_fill']
    current_row += 1
    
    summary_items = [
        ("Assessment Date:", datetime.now().strftime('%d.%m.%Y')),
        ("Overall Compliance %:", "=Summary_Dashboard!G11"),
        ("Total Requirements:", "=Summary_Dashboard!B11"),
        ("Critical Gaps:", "=Summary_Dashboard!E11"),
    ]
    
    for label, value in summary_items:
        current_row = add_metadata_row(ws, current_row, label, value, styles)
    
    current_row += 1
    
    # Approvers Table
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "Required Approvals"
    cell.font = styles['subheader_font']
    cell.fill = styles['subheader_fill']
    current_row += 1
    
    # Table headers
    ws[f'A{current_row}'] = "Role"
    ws[f'B{current_row}'] = "Details"
    ws[f'A{current_row}'].font = styles['column_header_font']
    ws[f'B{current_row}'].font = styles['column_header_font']
    ws[f'A{current_row}'].fill = styles['column_header_fill']
    ws[f'B{current_row}'].fill = styles['column_header_fill']
    current_row += 1
    
    # Approver roles
    approvers = [
        "Application Security Lead (Technical Review)",
        "Chief Information Security Officer (CISO Approval)",
        "Legal Counsel (License Compliance Approval)",
        "Procurement Lead (Vendor Management Approval)",
    ]
    
    for approver_role in approvers:
        ws[f'A{current_row}'] = approver_role
        ws[f'A{current_row}'].font = styles['bold_font']
        ws[f'A{current_row}'].alignment = styles['align_left']
        current_row += 1
        
        # Details fields
        fields = [
            ("Name:", ""),
            ("Signature:", ""),
            ("Date:", ""),
            ("Decision:", ""),
            ("Comments:", ""),
        ]
        
        for field_label, field_value in fields:
            ws[f'A{current_row}'] = f"  {field_label}"
            ws[f'B{current_row}'] = field_value
            ws[f'B{current_row}'].fill = styles['input_fill']
            
            if field_label == "Decision:":
                validations['approval_decision'].add(f'B{current_row}')
            
            current_row += 1
        
        current_row += 1  # Space between approvers
    
    print("✓ Approval_Sign_Off sheet created")


# ============================================================================
# SECTION 6: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function to generate the assessment workbook."""
    print("=" * 70)
    print("ISMS Control A.8.28.4 - Third-Party & OSS Assessment Generator")
    print("ISO/IEC 27001:2022 Control A.8.28 (Secure Coding)")
    print("=" * 70)
    print()
    
    # Initialize workbook
    print("Initializing workbook...")
    wb = create_workbook()
    styles = get_style_definitions()
    validations = create_data_validations()
    column_widths = get_column_widths()
    print("✓ Workbook initialized")
    print()
    
    # Create Instructions sheet
    print("Creating Instructions sheet...")
    create_instructions_sheet(wb, styles)
    print()
    
    # Create domain assessment sheets
    print("Creating domain assessment sheets...")
    
    create_domain_assessment_sheet(
        wb, 
        "Vendor_Security_Assessment",
        "Domain 1: Vendor Security Assessment",
        get_vendor_security_requirements(),
        styles,
        validations,
        column_widths
    )
    
    create_domain_assessment_sheet(
        wb,
        "OSS_Management",
        "Domain 2: Open Source Software Management",
        get_oss_management_requirements(),
        styles,
        validations,
        column_widths
    )
    
    create_domain_assessment_sheet(
        wb,
        "Dependency_Vulnerability_Mgmt",
        "Domain 3: Dependency Vulnerability Management",
        get_dependency_vulnerability_requirements(),
        styles,
        validations,
        column_widths
    )
    
    create_domain_assessment_sheet(
        wb,
        "Third_Party_Code_Review",
        "Domain 4: Third-Party Code Review & Integration",
        get_third_party_code_review_requirements(),
        styles,
        validations,
        column_widths
    )
    
    create_domain_assessment_sheet(
        wb,
        "License_Compliance",
        "Domain 5: License Compliance & Legal Risk",
        get_license_compliance_requirements(),
        styles,
        validations,
        column_widths
    )
    
    print()
    
    # Create supporting sheets
    print("Creating supporting sheets...")
    create_summary_dashboard(wb, styles)
    create_evidence_register(wb, styles, validations)
    create_gap_analysis(wb, styles, validations)
    create_approval_signoff(wb, styles, validations)
    print()
    
    # Save workbook
    print()
    print("💾 Saving workbook...")
    filename = f"ISMS-IMP-A.8.28.4_Third_Party_OSS_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    # Summary
    print()
    print("=" * 70)
    print("\u2705 SUCCESS: Third-Party & OSS Assessment workbook generated!")
    print("=" * 70)
    print()
    print(f"📁 File: {filename}")
    print(f"📊 Total Requirements: 90 (18 per domain × 5 domains)")
    print()
    print("Workbook Contents:")
    print("  1. Instructions - Assessment guidance and methodology")
    print("  2. Vendor_Security_Assessment - 18 requirements")
    print("  3. OSS_Management - 18 requirements")
    print("  4. Dependency_Vulnerability_Mgmt - 18 requirements")
    print("  5. Third_Party_Code_Review - 18 requirements")
    print("  6. License_Compliance - 18 requirements")
    print("  7. Summary_Dashboard - Compliance metrics")
    print("  8. Evidence_Register - Evidence tracking")
    print("  9. Gap_Analysis - Remediation tracking")
    print(" 10. Approval_Sign_Off - Formal approval")
    print()
    print("Total Requirements: 90 (18 per domain)")
    print()
    print("Next Steps:")
    print("  1. Distribute workbook to assessment team")
    print("  2. Complete domain assessments")
    print("  3. Collect and register evidence")
    print("  4. Track gaps and remediation")
    print("  5. Obtain formal approvals")
    print()
    print("Remember: Focus on what WORKS, not what's DOCUMENTED!")
    print("'The first principle is that you must not fool yourself' - Feynman")
    print("=" * 70)
    
    return 0


if __name__ == "__main__":
    try:
        output_file = main()
        sys.exit(0)
    except Exception as e:
        print(f"\n\u274C ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)