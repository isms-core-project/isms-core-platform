#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.13.4 - BC/DR Testing Results Tracker Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.13 (Backup), A.8.14 (Redundancy), A.5.30 (ICT BC)
Assessment Domain 4 of 4: Recovery Testing and Validation

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific BC/DR testing strategy, testing frequency requirements,
and validation procedures.

Key customization areas:
1. Testing frequency requirements per system tier
2. Test types and scope definitions (restore, failover, full DR scenario)
3. Success criteria and acceptance thresholds
4. Overdue testing alert thresholds
5. Testing compliance scoring methodology

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.13/14/5.30 BC/DR Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for tracking
and validating BC/DR testing activities across backup restore tests, failover
tests, and full disaster recovery scenario exercises.

**Purpose:**
Enables systematic tracking of BC/DR testing against ISO 27001:2022 Controls
A.8.13, A.8.14, and A.5.30 requirements, supporting the critical principle:
"Untested Recovery = No Recovery"

**Assessment Scope:**
- System inventory requiring regular testing
- Backup restore test tracking (per system/dataset)
- Failover test tracking (per redundant system)
- Full DR scenario test tracking (end-to-end exercises)
- Test frequency compliance monitoring
- Test results documentation (success/partial/failure)
- Issues identified during testing
- Remediation tracking for test failures
- Testing trend analysis (improving vs. degrading)
- Gap analysis for untested systems
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and testing methodology
2. Testing Results Tracker - Comprehensive test inventory and results (110 rows: 10 examples + 100 data entry)
3. Evidence Register - Audit evidence tracking and documentation (100 rows)
4. Approval & Sign-Off - Multi-level stakeholder review and approval workflow

**Key Features:**
- Data validation with comprehensive dropdown lists (test types, results, frequencies)
- Conditional formatting for overdue testing and failed tests
- Automated testing compliance calculations (last test vs. required frequency)
- Automated overdue test identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- 3-level approval workflow (Assessor -> ISO Officer -> CISO)
- Exception handling with graceful error reporting
- Professional styling without Excel repair warnings

**Integration:**
This assessment feeds into the ISMS-IMP-A.8.13.5 BC/DR Compliance Dashboard,
which consolidates data from all four BC/DR assessment domains for executive
oversight and audit readiness. Testing results validate capabilities assessed
in Domains 1 (Backup) and 2 (Redundancy).

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl --break-system-packages

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a813_4_testing_results.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a813_4_testing_results.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a813_4_testing_results.py --date 20250125

Output:
    File: ISMS_Assessment_Testing_Results.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize testing frequency requirements per system tier
    2. Inventory all systems requiring regular BC/DR testing
    3. Document backup restore test results
    4. Document failover test results
    5. Document full DR scenario test results
    6. Analyze testing compliance (on schedule vs. overdue)
    7. Track remediation for failed tests or identified issues
    8. Define testing schedule for untested systems
    9. Obtain stakeholder approvals
    10. Feed results into ISMS-IMP-A.8.13.5 BC/DR Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.13, A.8.14, A.5.30
Assessment Domain:    4 of 4 (BC/DR Testing & Validation)
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.13-14-5.30: BC/DR Framework Policy (Governance)
    - ISMS-POL-A.8.13-14-5.30-S5: Testing Methodology and Evidence Framework
    - ISMS-IMP-A.8.13-14-5.30-S4: Recovery Testing Process
    - ISMS-IMP-A.8.13-14-5.30-S5: BC/DR Assessment Guide
    - ISMS-IMP-A.8.13.1: Backup Inventory Assessment (Domain 1)
    - ISMS-IMP-A.8.13.2: Redundancy Analysis Assessment (Domain 2)
    - ISMS-IMP-A.8.13.3: RPO/RTO Compliance Matrix (Domain 3)
    - ISMS-IMP-A.8.13.5: BC/DR Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.13-14-5.30-S5 specification
    - Supports comprehensive BC/DR testing tracking and validation
    - Integrated with ISMS-IMP-A.8.13.5 BC/DR Compliance Dashboard
    - Includes automated testing compliance monitoring

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Untested Recovery = No Recovery:**
This assessment embodies the most critical BC/DR principle: recovery capabilities
are meaningless without regular testing. Systems without recent successful tests
should be treated as UNRECOVERABLE regardless of backup or redundancy architecture.

**Testing Frequency Requirements:**
Testing frequency should be risk-based and aligned with system criticality:
- Tier 1 (Critical): Quarterly backup restore tests, semi-annual failover tests
- Tier 2 (Important): Semi-annual backup restore tests, annual failover tests
- Tier 3 (Standard): Annual backup restore tests
- Tier 4 (Low): Opportunistic testing or test-on-change

Customize based on your risk tolerance and regulatory requirements.

**Test Types:**
Different test types validate different capabilities:
- **Backup Restore Test**: Validates backup integrity and restore procedures (A.8.13)
- **Failover Test**: Validates redundancy and failover mechanisms (A.8.14)
- **Full DR Scenario**: Validates end-to-end recovery including people, process, technology (A.5.30)

All three are required for comprehensive BC/DR validation.

**Test Results Interpretation:**
- **Success**: Test completed within RTO, all functionality verified
- **Partial Success**: Test completed but issues identified (acceptable if documented)
- **Failure**: Test did not complete or major functionality unavailable
- **Not Tested**: HIGH RISK - assume recovery will fail

**Testing Without Production Impact:**
BC/DR testing should not disrupt production systems:
- Restore to isolated test environment (not production)
- Failover during maintenance windows
- Use snapshots or clones for testing

Document testing methodology to demonstrate safe practices.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Evidence of regular testing (test reports, screenshots, logs)
- Documented testing procedures
- Remediation tracking for failed tests
- Management approval of testing schedule

**Data Protection:**
Assessment workbooks contain sensitive operational information including:
- Testing schedules (security-sensitive timing information)
- Test failures and issues identified
- Recovery procedures and timings

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Monthly: Update with latest test results
- Quarterly: Analyze testing trends and compliance rates
- Semi-annually: Review testing frequency requirements
- Annually: Complete full reassessment of testing strategy
- Ad-hoc: After failed tests or recovery incidents

**Quality Assurance:**
Have BC/DR coordinators, system administrators, and business continuity managers
validate test results before using for compliance reporting.

**Regulatory Alignment:**
Ensure testing frequency aligns with applicable regulatory requirements:
- Finance: DORA requires regular testing of ICT systems
- Healthcare: HIPAA requires contingency plan testing and revision
- Critical Infrastructure: Sector-specific testing mandates

**Learning from Test Failures:**
Test failures are GOOD - they identify issues before real disasters:
- Document all issues found during testing
- Track remediation with priority and timelines
- Re-test after remediation
- Update recovery procedures based on lessons learned

The goal of testing is to find and fix problems, not to achieve 100% success rate.

**Testing vs. Real Incidents:**
Testing validates capability but doesn't guarantee success during real incidents:
- Tests are controlled and planned
- Real incidents involve stress, time pressure, and unknowns
- Both testing and incident response experience are valuable

Document both test results AND actual recovery incidents.

================================================================================
"""

import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================================
# CONFIGURATION
# ============================================================================

WORKBOOK_TITLE = "BC/DR Testing Results Tracker"
VERSION = "1.0"
CONTROLS = "A.8.13 (Information Backup)"


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
DISK = '\U0001F4BE'   # 💾 Floppy Disk
GLOBE = '\U0001F310'  # 🌐 Globe
SEARCH = '\U0001F50D' # 🔍 Magnifying Glass
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

DOCUMENT_ID = "ISMS-IMP-A.8.13.4"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Color scheme (consistent with reference implementations)
HEADER_FILL = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
COLUMN_HEADER_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

# Fonts
HEADER_FONT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
SUBHEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
BOLD_FONT = Font(name='Calibri', size=10, bold=True)
NORMAL_FONT = Font(name='Calibri', size=10)

# Borders
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def apply_style(cell, font=None, fill=None, alignment=None, border=None):
    """Apply multiple styles to a cell"""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border

def create_base_validations(ws):
    """
    Create comprehensive data validation objects for dropdowns.
    
    Returns dict of validation objects to be applied to cells.
    All validations are added to worksheet once, then applied to multiple cells.
    """
    validations = {
        'test_type': DataValidation(
            type="list",
            formula1='"Backup Restore,Failover Test,Full DR Scenario,Tabletop Exercise,Component Test,Integration Test"',
            allow_blank=False
        ),
        'test_result': DataValidation(
            type="list",
            formula1=f'"{CHECK} Success,⚠️ Partial Success,❌ Failure,⏳ In Progress,➖ Not Started,🔄 Cancelled"',
            allow_blank=False
        ),
        'criticality': DataValidation(
            type="list",
            formula1='"Tier 1 - Critical,Tier 2 - Important,Tier 3 - Standard,Tier 4 - Low"',
            allow_blank=False
        ),
        'test_frequency': DataValidation(
            type="list",
            formula1='"Quarterly,Semi-Annual,Annual,Ad-Hoc,Not Required"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Compliant,⚠️ Overdue,❌ Non-Compliant,❓ Unknown"',
            allow_blank=False
        ),
        'issue_severity': DataValidation(
            type="list",
            formula1='"🔴 Critical,🟡 High,🟢 Medium,⚪ Low"',
            allow_blank=False
        ),
        'issue_status': DataValidation(
            type="list",
            formula1='"🔴 Open,⏳ In Progress,✅ Closed,➖ Risk Accepted"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Test Report,Screenshot,Log File,Video Recording,Configuration File,Runbook,Issue Ticket,Sign-Off,Other"',
            allow_blank=False
        ),
        'verification_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Verified,⏳ Pending,❌ Not Verified"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1='"Draft,Under Review,Final,Requires Remediation"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected,Requires Rework"',
            allow_blank=False
        ),
        'recommendation': DataValidation(
            type="list",
            formula1='"Approve,Approve with Conditions,Reject,Require Rework"',
            allow_blank=False
        ),
    }
    
    # Add all validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    return validations

# ============================================================================
# WORKSHEET: INSTRUCTIONS
# ============================================================================

def create_instructions_sheet(wb):
    """Create comprehensive instructions worksheet"""
    ws = wb.create_sheet(title="Instructions", index=0)
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f'ISMS BC/DR Assessment - {WORKBOOK_TITLE}'
    apply_style(ws['A1'], font=Font(name='Calibri', size=16, bold=True, color='003366'),
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35
    
    # Document metadata
    ws['A3'] = 'Document ID:'
    ws['A3'].font = Font(bold=True)
    ws['B3'] = DOCUMENT_ID
    
    ws['A4'] = 'Assessment:'
    ws['A4'].font = Font(bold=True)
    ws['B4'] = 'BC/DR Testing Results'
    
    ws['A5'] = 'Version:'
    ws['A5'].font = Font(bold=True)
    ws['B5'] = VERSION
    
    ws['A6'] = 'Generated:'
    ws['A6'].font = Font(bold=True)
    ws['B6'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    ws.column_dimensions['B'].width = 40
    
    # Purpose
    row = 8
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'PURPOSE'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    purpose_text = [
        'This workbook tracks all BC/DR testing activities, results, issues discovered, and testing compliance.',
        'Use this to demonstrate that backup, failover, and DR capabilities are regularly TESTED, not just assumed.',
        '',
        'KEY PRINCIPLE: "Test or it doesn\'t work" - Untested recovery capability = no recovery capability.',
    ]
    
    for text in purpose_text:
        ws[f'A{row}'] = text
        ws[f'A{row}'].font = NORMAL_FONT
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Workflow
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'TESTING WORKFLOW'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    workflow_steps = [
        '1. Plan tests in Test_Schedule (define annual test plan)',
        '2. Execute tests according to schedule',
        '3. Log results in Test_Results_Log (document actual test execution)',
        '4. Document issues in Issue_Remediation (problems discovered during testing)',
        '5. Track remediation of issues to closure',
        '6. Review Testing_Compliance (annual compliance assessment)',
        '7. Collect evidence in Evidence_Register (MINIMUM 5 evidence items for audit)',
        '8. Review Summary dashboard for overall testing compliance',
        '9. Complete Approval_Sign_Off workflow (3 levels: Assessor → ISO → CISO)',
        '10. Repeat annually for continuous compliance',
    ]
    
    for step in workflow_steps:
        ws[f'A{row}'] = step
        ws[f'A{row}'].font = NORMAL_FONT
        row += 1
    
    # Test Types
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'TEST TYPES EXPLAINED'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    test_types = [
        ('Backup Restore:', 'Test restoring data from backup - verify RPO capability (A.8.13)'),
        ('Failover Test:', 'Test switching to redundant system - verify RTO capability (A.8.14)'),
        ('Full DR Scenario:', 'End-to-end disaster recovery test - complete system recovery (A.5.30)'),
        ('Tabletop Exercise:', 'Discussion-based walkthrough of DR procedures (no actual execution)'),
        ('Component Test:', 'Test individual component (e.g., database restore only)'),
        ('Integration Test:', 'Test multiple components together (e.g., app + database recovery)'),
    ]
    
    for label, description in test_types:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = description
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'B{row}'].font = NORMAL_FONT
        row += 1
    
    # Worksheets Description
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'WORKSHEET DESCRIPTIONS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    worksheet_descriptions = [
        ('Instructions:', 'This sheet - comprehensive usage guide'),
        ('Summary:', 'Executive dashboard with testing compliance metrics'),
        ('Test_Schedule:', 'Annual test plan - what tests to perform when (110 rows)'),
        ('Test_Results_Log:', 'All test executions - actual test results (110 rows)'),
        ('Issue_Remediation:', 'Issues discovered during testing - track to closure (110 rows)'),
        ('Testing_Compliance:', 'Annual compliance assessment - are we testing enough? (110 rows)'),
        ('Evidence_Register:', '100 evidence slots for audit traceability (minimum 5 required)'),
        ('Approval_Sign_Off:', '3-level approval workflow (Assessor → ISO → CISO)'),
    ]
    
    for label, description in worksheet_descriptions:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = description
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'B{row}'].font = NORMAL_FONT
        row += 1
    
    # Testing Frequency Guidelines
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'TESTING FREQUENCY GUIDELINES'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    frequency_guidelines = [
        'TIER 1 - CRITICAL SYSTEMS:',
        '  • Backup Restore: Quarterly (every 3 months)',
        '  • Failover Test: Quarterly (every 3 months)',
        '  • Full DR Scenario: Annual (at least once per year)',
        '',
        'TIER 2 - IMPORTANT SYSTEMS:',
        '  • Backup Restore: Semi-Annual (every 6 months)',
        '  • Failover Test: Semi-Annual (every 6 months)',
        '  • Full DR Scenario: Annual (at least once per year)',
        '',
        'TIER 3 - STANDARD SYSTEMS:',
        '  • Backup Restore: Annual (at least once per year)',
        '  • Failover Test: Annual (if redundancy exists)',
        '',
        'TIER 4 - LOW SYSTEMS:',
        '  • Testing may be ad-hoc or not required (risk acceptance)',
    ]
    
    for guideline in frequency_guidelines:
        ws[f'A{row}'] = guideline
        if 'TIER' in guideline:
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        else:
            ws[f'A{row}'].font = NORMAL_FONT
        row += 1
    
    # Status Indicators Legend
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'STATUS INDICATOR LEGEND'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    legend_items = [
        (f'{CHECK}', 'Success / Compliant / Closed / Verified'),
        (f'{XMARK}', 'Failure / Non-Compliant / Open'),
        (f'{WARNING}', 'Partial Success / Overdue / Warning'),
        ('⏳', 'In Progress / Pending'),
        ('❓', 'Unknown / Not Tested'),
        ('➖', 'Not Started / Risk Accepted / Not Applicable'),
        ('🔄', 'Cancelled / Rescheduled'),
        ('🔴', 'Critical Severity / Open Issue'),
        ('🟡', 'High Severity'),
        ('🟢', 'Medium Severity'),
        ('⚪', 'Low Severity'),
    ]
    
    for emoji, description in legend_items:
        ws[f'A{row}'] = emoji
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'B{row}'] = description
        row += 1
    
    # Critical Notes
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'CRITICAL TESTING REQUIREMENTS'
    apply_style(ws[f'A{row}'], font=Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
                fill=PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'),
                alignment=Alignment(horizontal='center'))
    
    row += 1
    critical_notes = [
        '❗ CRITICAL SYSTEMS: Tier 1 systems MUST be tested quarterly (backup & failover)',
        '❗ EVIDENCE: Minimum 5 evidence items required (test reports, logs, screenshots)',
        '❗ ISSUES: All 🔴 Critical issues must have remediation plans',
        '❗ APPROVAL: All 3 sign-off levels required (Assessor, ISO, CISO)',
        '❗ ANNUAL DR: Full DR scenario test required at least once per year',
        '❗ DOCUMENTATION: All tests must be documented with results (success/failure)',
    ]
    
    for note in critical_notes:
        ws[f'A{row}'] = note
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color='C00000')
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 50
    for col in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15
    
    return ws

# ============================================================================
# WORKSHEET: SUMMARY DASHBOARD
# ============================================================================

def create_summary_sheet(wb):
    """Create Summary dashboard worksheet (placed as 2nd sheet)"""
    ws = wb.create_sheet(title="Summary", index=1)
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = 'BC/DR TESTING COMPLIANCE DASHBOARD'
    apply_style(ws['A1'], font=Font(name='Calibri', size=16, bold=True, color='003366'),
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f'Assessment Date: {datetime.now().strftime("%Y-%m-%d")} | Assessment ID: {DOCUMENT_ID}'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(italic=True, size=11)
    
    # Testing Activity Summary
    row = 4
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'TESTING ACTIVITY SUMMARY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    activity_metrics = [
        ('Total Tests Scheduled:', '=COUNTA(Test_Schedule!A5:A114)'),
        ('Tests Completed (✅ Success):', '=COUNTIF(Test_Results_Log!F5:F114,f"{CHECK}*")'),
        ('Tests Partial Success (⚠️):', '=COUNTIF(Test_Results_Log!F5:F114,f"{WARNING}*")'),
        ('Tests Failed (❌):', '=COUNTIF(Test_Results_Log!F5:F114,f"{XMARK}*")'),
        ('Tests In Progress (⏳):', '=COUNTIF(Test_Results_Log!F5:F114,"⏳*")'),
        ('Tests Not Started (➖):', '=COUNTIF(Test_Results_Log!F5:F114,"➖*")'),
        ('', ''),
        ('Test Success Rate:', '=IF(B5>0,(B6+B7*0.5)/B5,0)'),
    ]
    
    for label, formula in activity_metrics:
        if label:
            ws[f'A{row}'] = label
            apply_style(ws[f'A{row}'], font=BOLD_FONT)
            ws[f'B{row}'] = formula
            apply_style(ws[f'B{row}'], font=Font(bold=True, size=12))
            if 'Rate' in label:
                ws[f'B{row}'].number_format = '0.0%'
        row += 1
    
    # Testing Compliance by Criticality
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'TESTING COMPLIANCE BY CRITICALITY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    criticality_metrics = [
        ('Tier 1 - Critical Systems:', '=COUNTIF(Testing_Compliance!C5:C114,"Tier 1*")'),
        ('Tier 1 - Compliant:', '=SUMPRODUCT((Testing_Compliance!C5:C114="Tier 1 - Critical")*(Testing_Compliance!G5:G114=f"{CHECK} Compliant"))'),
        ('Tier 1 - Compliance Rate:', '=IF(B15>0,B16/B15,0)'),
        ('', ''),
        ('Tier 2 - Important Systems:', '=COUNTIF(Testing_Compliance!C5:C114,"Tier 2*")'),
        ('Tier 2 - Compliant:', '=SUMPRODUCT((Testing_Compliance!C5:C114="Tier 2 - Important")*(Testing_Compliance!G5:G114=f"{CHECK} Compliant"))'),
        ('Tier 2 - Compliance Rate:', '=IF(B19>0,B20/B19,0)'),
    ]
    
    for label, formula in criticality_metrics:
        if label:
            ws[f'A{row}'] = label
            apply_style(ws[f'A{row}'], font=BOLD_FONT)
            ws[f'B{row}'] = formula
            apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
            if 'Rate' in label:
                ws[f'B{row}'].number_format = '0.0%'
        row += 1
    
    # Issue Summary
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'ISSUE SUMMARY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    issue_metrics = [
        ('Total Issues Identified:', '=COUNTA(Issue_Remediation!A5:A114)'),
        ('🔴 Critical Issues:', '=COUNTIF(Issue_Remediation!E5:E114,"🔴*")'),
        ('🟡 High Severity Issues:', '=COUNTIF(Issue_Remediation!E5:E114,"🟡*")'),
        ('🟢 Medium Severity Issues:', '=COUNTIF(Issue_Remediation!E5:E114,"🟢*")'),
        ('⚪ Low Severity Issues:', '=COUNTIF(Issue_Remediation!E5:E114,"⚪*")'),
        ('', ''),
        ('Open Issues (🔴):', '=COUNTIF(Issue_Remediation!F5:F114,"🔴*")'),
        ('In Progress (⏳):', '=COUNTIF(Issue_Remediation!F5:F114,"⏳*")'),
        ('Closed Issues (✅):', '=COUNTIF(Issue_Remediation!F5:F114,f"{CHECK}*")'),
    ]
    
    for label, formula in issue_metrics:
        if label:
            ws[f'A{row}'] = label
            apply_style(ws[f'A{row}'], font=BOLD_FONT)
            ws[f'B{row}'] = formula
            apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1
    
    # Evidence & Approval Status
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'EVIDENCE & APPROVAL STATUS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    evidence_metrics = [
        ('Evidence Items Collected:', '=COUNTA(Evidence_Register!A5:A104)'),
        ('Verified Evidence:', '=COUNTIF(Evidence_Register!H5:H104,f"{CHECK}*")'),
        ('Minimum Evidence Required:', '5'),
        ('Evidence Compliance:', '=IF(B39>=B41,f"{CHECK} Sufficient",f"{XMARK} Insufficient")'),
        ('', ''),
        ('Assessment Status:', '=Approval_Sign_Off!B14'),
        ('Level 1 - Assessor Completed:', '=IF(Approval_Sign_Off!B26<>"",f"{CHECK} Complete","⏳ Pending")'),
        ('Level 2 - ISO Review:', '=IF(Approval_Sign_Off!B37<>"",f"{CHECK} Complete","⏳ Pending")'),
        ('Level 3 - CISO Approval:', '=IF(Approval_Sign_Off!B49<>"",f"{CHECK} Complete","⏳ Pending")'),
    ]
    
    for label, formula in evidence_metrics:
        if label:
            ws[f'A{row}'] = label
            apply_style(ws[f'A{row}'], font=BOLD_FONT)
            ws[f'B{row}'] = formula
            if label == 'Minimum Evidence Required:':
                apply_style(ws[f'B{row}'], font=Font(bold=True, size=11, color='C00000'))
            else:
                apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1
    
    # Priority Actions
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'PRIORITY ACTIONS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    priority_actions = [
        '1. Address all 🔴 Critical issues discovered during testing',
        '2. Complete all overdue tests (⚠️ Overdue in Testing_Compliance)',
        '3. Schedule quarterly tests for Tier 1 systems',
        '4. Remediate test failures (❌ Failed tests in Test_Results_Log)',
        '5. Document lessons learned from partial successes (⚠️)',
        '6. Collect minimum 5 evidence items in Evidence_Register',
        '7. Complete all 3 approval levels (Assessor → ISO → CISO)',
        '8. Plan next annual full DR scenario test',
        '9. Update runbooks based on test findings',
    ]
    
    for action in priority_actions:
        ws[f'A{row}'] = action
        ws[f'A{row}'].font = NORMAL_FONT
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 25
    for col in ['C', 'D', 'E']:
        ws.column_dimensions[col].width = 15
    
    return ws

# ============================================================================
# WORKSHEET: TEST SCHEDULE
# ============================================================================

def create_test_schedule_sheet(wb):
    """Create Test Schedule worksheet (annual test planning)"""
    ws = wb.create_sheet(title="Test_Schedule")
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = 'BC/DR TEST SCHEDULE (Annual Plan)'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:H2')
    ws['A2'] = 'Define annual test plan - what tests to perform when (schedule all required tests)'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'Test ID',
        'System Name',
        'Test Type',
        'Criticality',
        'Required Frequency',
        'Scheduled Date',
        'Test Owner',
        'Notes / Prerequisites'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    # Apply data validations and styling for 110 rows
    row = 5
    for i in range(110):
        current_row = row + i
        
        # Test ID (auto-generated TEST-001 to TEST-110)
        ws[f'A{current_row}'] = f'TEST-{i+1:03d}'
        apply_style(ws[f'A{current_row}'], font=Font(bold=True, size=9), border=THIN_BORDER)
        
        # System Name, Test Owner, Notes (user input)
        for col in ['B', 'G', 'H']:
            if i >= 10:
                apply_style(ws[f'{col}{current_row}'], fill=INPUT_FILL, border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
            else:
                apply_style(ws[f'{col}{current_row}'], border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
        
        # Test Type dropdown
        validations['test_type'].add(f'C{current_row}')
        apply_style(ws[f'C{current_row}'], border=THIN_BORDER)
        
        # Criticality dropdown
        validations['criticality'].add(f'D{current_row}')
        apply_style(ws[f'D{current_row}'], border=THIN_BORDER)
        
        # Required Frequency dropdown
        validations['test_frequency'].add(f'E{current_row}')
        apply_style(ws[f'E{current_row}'], border=THIN_BORDER)
        
        # Scheduled Date (user input, date format)
        if i >= 10:
            apply_style(ws[f'F{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)
        else:
            apply_style(ws[f'F{current_row}'], border=THIN_BORDER)
        ws[f'F{current_row}'].number_format = 'YYYY-MM-DD'
    
    # Example test schedule (10 realistic examples)
    examples = [
        ['TEST-001', 'E-Commerce Website', 'Backup Restore', 'Tier 1 - Critical', 'Quarterly',
         (datetime.now() + timedelta(days=15)).strftime('%Y-%m-%d'), 'Infrastructure Team',
         'Restore full website DB, verify data integrity, test 4hr restore time'],
        ['TEST-002', 'E-Commerce Website', 'Failover Test', 'Tier 1 - Critical', 'Quarterly',
         (datetime.now() + timedelta(days=45)).strftime('%Y-%m-%d'), 'Infrastructure Team',
         'Test Active-Passive cluster failover, verify <30min RTO'],
        ['TEST-003', 'Payment Gateway', 'Failover Test', 'Tier 1 - Critical', 'Quarterly',
         (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d'), 'Payments Team',
         'Test Active-Active failover, critical system, must coordinate with processor'],
        ['TEST-004', 'Email System', 'Backup Restore', 'Tier 1 - Critical', 'Quarterly',
         (datetime.now() + timedelta(days=60)).strftime('%Y-%m-%d'), 'IT Operations',
         'Restore mailbox data, test 12hr restore time, verify user access'],
        ['TEST-005', 'CRM System', 'Backup Restore', 'Tier 1 - Critical', 'Quarterly',
         (datetime.now() + timedelta(days=75)).strftime('%Y-%m-%d'), 'App Support Team',
         'Restore CRM DB, verify integrations work, test 8hr restore time'],
        ['TEST-006', 'ERP System', 'Backup Restore', 'Tier 2 - Important', 'Semi-Annual',
         (datetime.now() + timedelta(days=90)).strftime('%Y-%m-%d'), 'ERP Team',
         'Restore ERP DB, complex dependencies, test 8hr restore, schedule during low activity'],
        ['TEST-007', 'HR Information System', 'Backup Restore', 'Tier 2 - Important', 'Semi-Annual',
         (datetime.now() + timedelta(days=120)).strftime('%Y-%m-%d'), 'HR IT Team',
         'Restore HR DB, sensitive data, coordinate with CISO, test 20hr restore'],
        ['TEST-008', 'All Critical Systems', 'Full DR Scenario', 'Tier 1 - Critical', 'Annual',
         (datetime.now() + timedelta(days=180)).strftime('%Y-%m-%d'), 'BC/DR Coordinator',
         'Full DR exercise - simulate datacenter loss, recover all Tier 1 systems'],
        ['TEST-009', 'File Server', 'Backup Restore', 'Tier 3 - Standard', 'Annual',
         (datetime.now() + timedelta(days=210)).strftime('%Y-%m-%d'), 'IT Operations',
         'Restore file share, test 24hr restore time, low priority'],
        ['TEST-010', 'BC/DR Team', 'Tabletop Exercise', 'Tier 1 - Critical', 'Annual',
         (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d'), 'BC/DR Coordinator',
         'Walkthrough DR procedures with stakeholders, no actual execution'],
    ]
    
    row = 5
    for idx, test_data in enumerate(examples, start=row):
        for col_idx, value in enumerate(test_data, start=1):
            if col_idx != 1:  # Skip Test ID (already set)
                cell = ws.cell(row=idx, column=col_idx+1 if col_idx > 1 else col_idx, value=value)
                apply_style(cell, border=THIN_BORDER)
                if col_idx in [2, 7, 8]:  # System Name, Owner, Notes
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Summary metrics
    summary_row = 117
    ws.merge_cells(f'A{summary_row}:H{summary_row}')
    ws[f'A{summary_row}'] = 'TEST SCHEDULE SUMMARY'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    summary_row += 1
    metrics = [
        ('Total Tests Scheduled:', f'=COUNTA(B5:B114)'),
        ('Backup Restore Tests:', f'=COUNTIF(C5:C114,"Backup*")'),
        ('Failover Tests:', f'=COUNTIF(C5:C114,"Failover*")'),
        ('Full DR Scenarios:', f'=COUNTIF(C5:C114,"Full DR*")'),
        ('Tier 1 System Tests:', f'=COUNTIF(D5:D114,"Tier 1*")'),
        ('Quarterly Tests:', f'=COUNTIF(E5:E114,"Quarterly")'),
    ]
    
    for label, formula in metrics:
        ws[f'A{summary_row}'] = label
        apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
        ws[f'B{summary_row}'] = formula
        apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
        summary_row += 1
    
    # Column widths
    widths = [12, 25, 20, 18, 18, 15, 20, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'A5'
    
    return ws

# ============================================================================
# WORKSHEET: TEST RESULTS LOG
# ============================================================================

def create_test_results_log_sheet(wb):
    """Create Test Results Log worksheet (all test executions and outcomes)"""
    ws = wb.create_sheet(title="Test_Results_Log")
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = 'BC/DR TEST RESULTS LOG (Execution History)'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:J2')
    ws['A2'] = 'Document all test executions - actual test results (log every test performed)'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'Result ID',
        'Test ID (Schedule)',
        'System Name',
        'Test Type',
        'Test Date',
        'Result',
        'Actual Duration (hrs)',
        'Tested By',
        'Issues Found',
        'Detailed Notes'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    # Apply data validations and styling for 110 rows
    row = 5
    for i in range(110):
        current_row = row + i
        
        # Result ID (auto-generated RES-001 to RES-110)
        ws[f'A{current_row}'] = f'RES-{i+1:03d}'
        apply_style(ws[f'A{current_row}'], font=Font(bold=True, size=9), border=THIN_BORDER)
        
        # Test ID, System Name, Tested By, Issues Found, Notes (user input)
        for col in ['B', 'C', 'H', 'I', 'J']:
            if i >= 10:
                apply_style(ws[f'{col}{current_row}'], fill=INPUT_FILL, border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
            else:
                apply_style(ws[f'{col}{current_row}'], border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
        
        # Test Type dropdown
        validations['test_type'].add(f'D{current_row}')
        apply_style(ws[f'D{current_row}'], border=THIN_BORDER)
        
        # Test Date (user input, date format)
        if i >= 10:
            apply_style(ws[f'E{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)
        else:
            apply_style(ws[f'E{current_row}'], border=THIN_BORDER)
        ws[f'E{current_row}'].number_format = 'YYYY-MM-DD'
        
        # Result dropdown
        validations['test_result'].add(f'F{current_row}')
        apply_style(ws[f'F{current_row}'], border=THIN_BORDER)
        
        # Actual Duration (user input, numeric)
        if i >= 10:
            apply_style(ws[f'G{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)
        else:
            apply_style(ws[f'G{current_row}'], border=THIN_BORDER)
    
    # Example test results (10 realistic examples)
    examples = [
        ['RES-001', 'TEST-001', 'E-Commerce Website', 'Backup Restore',
         (datetime.now() - timedelta(days=5)).strftime('%Y-%m-%d'), f'{CHECK} Success', 3.8,
         'Infrastructure Team', 'None', 'Full restore successful, verified data integrity, 3.8hr restore time'],
        ['RES-002', 'TEST-002', 'E-Commerce Website', 'Failover Test',
         (datetime.now() - timedelta(days=15)).strftime('%Y-%m-%d'), f'{WARNING} Partial Success', 0.6,
         'Infrastructure Team', 'ISS-002, ISS-003',
         'Failover worked but took 36min (exceeds 30min RTO), found monitoring gaps'],
        ['RES-003', 'TEST-003', 'Payment Gateway', 'Failover Test',
         (datetime.now() - timedelta(days=20)).strftime('%Y-%m-%d'), f'{CHECK} Success', 0.08,
         'Payments Team', 'None', 'Failover successful, ~5min, no transaction loss, excellent result'],
        ['RES-004', 'TEST-004', 'Email System', 'Backup Restore',
         (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'), f'{XMARK} Failure', 16.0,
         'IT Operations', 'ISS-001',
         'Restore failed, backup corrupted, root cause: backup job misconfiguration'],
        ['RES-005', 'TEST-005', 'CRM System', 'Backup Restore',
         (datetime.now() - timedelta(days=45)).strftime('%Y-%m-%d'), f'{WARNING} Partial Success', 9.5,
         'App Support Team', 'ISS-004',
         'Restore successful but took 9.5hr (exceeds 8hr RTO), integration issues post-restore'],
        ['RES-006', 'TEST-006', 'ERP System', 'Backup Restore',
         (datetime.now() - timedelta(days=60)).strftime('%Y-%m-%d'), f'{CHECK} Success', 7.2,
         'ERP Team', 'None', 'Restore successful, complex dependencies resolved, 7.2hr within 8hr target'],
        ['RES-007', 'TEST-010', 'BC/DR Team', 'Tabletop Exercise',
         (datetime.now() - timedelta(days=10)).strftime('%Y-%m-%d'), f'{CHECK} Success', 2.0,
         'BC/DR Coordinator', 'ISS-005',
         'Tabletop successful, identified gaps in communication procedures'],
        ['RES-008', 'TEST-008', 'All Critical Systems', 'Full DR Scenario',
         (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d'), f'{WARNING} Partial Success', 12.0,
         'BC/DR Coordinator', 'ISS-006, ISS-007, ISS-008',
         'DR scenario partially successful, 3/5 Tier 1 systems recovered within RTO'],
        ['RES-009', 'TEST-009', 'File Server', 'Backup Restore',
         (datetime.now() - timedelta(days=120)).strftime('%Y-%m-%d'), f'{CHECK} Success', 22.0,
         'IT Operations', 'None', 'Restore successful, 22hr within 24hr target, Tier 3 system'],
        ['RES-010', 'TEST-007', 'HR Information System', 'Backup Restore',
         (datetime.now() - timedelta(days=135)).strftime('%Y-%m-%d'), f'{CHECK} Success', 18.5,
         'HR IT Team', 'None', 'Restore successful, 18.5hr within 20hr target, sensitive data verified'],
    ]
    
    row = 5
    for idx, result_data in enumerate(examples, start=row):
        for col_idx, value in enumerate(result_data, start=1):
            if col_idx != 1:  # Skip Result ID (already set)
                cell = ws.cell(row=idx, column=col_idx+1 if col_idx > 1 else col_idx, value=value)
                apply_style(cell, border=THIN_BORDER)
                if col_idx in [3, 8, 9, 10]:  # System Name, Tested By, Issues, Notes
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Summary metrics
    summary_row = 117
    ws.merge_cells(f'A{summary_row}:J{summary_row}')
    ws[f'A{summary_row}'] = 'TEST RESULTS SUMMARY'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    summary_row += 1
    metrics = [
        ('Total Tests Executed:', f'=COUNTA(C5:C114)'),
        (f'{CHECK} Successful Tests:', f'=COUNTIF(F5:F114,f"{CHECK}*")'),
        (f'{WARNING} Partial Success:', f'=COUNTIF(F5:F114,f"{WARNING}*")'),
        (f'{XMARK} Failed Tests:', f'=COUNTIF(F5:F114,f"{XMARK}*")'),
        ('Success Rate:', f'=IF(B{summary_row}>0,(B{summary_row+1}+B{summary_row+2}*0.5)/B{summary_row},0)'),
        ('Average Test Duration (hrs):', f'=IFERROR(AVERAGE(G5:G114),0)'),
    ]
    
    for label, formula in metrics:
        ws[f'A{summary_row}'] = label
        apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
        ws[f'B{summary_row}'] = formula
        apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
        if 'Rate' in label:
            ws[f'B{summary_row}'].number_format = '0.0%'
        elif 'Duration' in label:
            ws[f'B{summary_row}'].number_format = '0.0'
        summary_row += 1
    
    # Column widths
    widths = [12, 18, 25, 20, 15, 18, 18, 20, 20, 45]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'A5'
    
    return ws

# ============================================================================
# WORKSHEET: ISSUE REMEDIATION
# ============================================================================

def create_issue_remediation_sheet(wb):
    """Create Issue Remediation worksheet (issues discovered during testing)"""
    ws = wb.create_sheet(title="Issue_Remediation")
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = 'ISSUE REMEDIATION TRACKER'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:H2')
    ws['A2'] = 'Track issues discovered during testing to closure'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'Issue ID',
        'Related Test (Result ID)',
        'Issue Description',
        'Root Cause',
        'Severity',
        'Status',
        'Remediation Plan',
        'Target Resolution Date'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    # Apply data validations and styling for 110 rows
    row = 5
    for i in range(110):
        current_row = row + i
        
        # Issue ID (auto-generated ISS-001 to ISS-110)
        ws[f'A{current_row}'] = f'ISS-{i+1:03d}'
        apply_style(ws[f'A{current_row}'], font=Font(bold=True, size=9), border=THIN_BORDER)
        
        # Related Test, Description, Root Cause, Remediation Plan (user input)
        for col in ['B', 'C', 'D', 'G']:
            if i >= 10:
                apply_style(ws[f'{col}{current_row}'], fill=INPUT_FILL, border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
            else:
                apply_style(ws[f'{col}{current_row}'], border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
        
        # Severity dropdown
        validations['issue_severity'].add(f'E{current_row}')
        apply_style(ws[f'E{current_row}'], border=THIN_BORDER)
        
        # Status dropdown
        validations['issue_status'].add(f'F{current_row}')
        apply_style(ws[f'F{current_row}'], border=THIN_BORDER)
        
        # Target Resolution Date (user input, date format)
        if i >= 10:
            apply_style(ws[f'H{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)
        else:
            apply_style(ws[f'H{current_row}'], border=THIN_BORDER)
        ws[f'H{current_row}'].number_format = 'YYYY-MM-DD'
    
    # Example issues (10 realistic issues from testing)
    examples = [
        ['ISS-001', 'RES-004', 'Email system backup corrupted - restore failed',
         'Backup job misconfigured, missing transaction log backups',
         '🔴 Critical', f'{CHECK} Closed',
         'Reconfigured backup job, added transaction log backups, verified with successful test',
         (datetime.now() - timedelta(days=20)).strftime('%Y-%m-%d')],
        ['ISS-002', 'RES-002', 'E-Commerce failover took 36min (exceeds 30min RTO)',
         'Manual DNS update required, not automated',
         '🟡 High', f'{CHECK} Closed',
         'Automated DNS failover, reduced to <30min, retested successfully',
         (datetime.now() - timedelta(days=10)).strftime('%Y-%m-%d')],
        ['ISS-003', 'RES-002', 'Monitoring did not alert during failover',
         'Monitoring not configured for cluster failover events',
         '🟢 Medium', f'{CHECK} Closed',
         'Configured monitoring for failover events, tested alerts',
         (datetime.now() - timedelta(days=8)).strftime('%Y-%m-%d')],
        ['ISS-004', 'RES-005', 'CRM restore took 9.5hr (exceeds 8hr RTO)',
         'Integration dependencies not documented in restore procedure',
         '🟡 High', '⏳ In Progress',
         'Update restore runbook with integration steps, optimize restore process',
         (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')],
        ['ISS-005', 'RES-007', 'Communication procedures gap identified in tabletop',
         'No clear escalation path defined for DR coordinator',
         '🟡 High', f'{CHECK} Closed',
         'Documented escalation procedures, updated BC/DR plan',
         (datetime.now() - timedelta(days=5)).strftime('%Y-%m-%d')],
        ['ISS-006', 'RES-008', 'Payment Gateway not recovered during DR scenario',
         'Active-Active failover requires manual processor notification',
         '🔴 Critical', '⏳ In Progress',
         'Automate processor notification, create runbook, retest in next DR scenario',
         (datetime.now() + timedelta(days=60)).strftime('%Y-%m-%d')],
        ['ISS-007', 'RES-008', 'CRM not recovered within RTO during DR scenario',
         'Complex dependencies, restore order not documented',
         '🟡 High', '⏳ In Progress',
         'Document system dependency tree, define restore order, update runbook',
         (datetime.now() + timedelta(days=45)).strftime('%Y-%m-%d')],
        ['ISS-008', 'RES-008', 'Network connectivity issues during DR scenario',
         'DR site network not pre-configured, manual setup required',
         '🟡 High', '🔴 Open',
         'Pre-configure DR site network, automate VPN setup, document procedures',
         (datetime.now() + timedelta(days=90)).strftime('%Y-%m-%d')],
        ['ISS-009', '[General]', 'Test documentation inconsistent across teams',
         'No standard template for test documentation',
         '🟢 Medium', f'{CHECK} Closed',
         'Created standard test report template, distributed to all teams',
         (datetime.now() - timedelta(days=15)).strftime('%Y-%m-%d')],
        ['ISS-010', '[General]', 'Lack of training for restore procedures',
         'Staff turnover, new team members not trained on DR procedures',
         '🟡 High', '⏳ In Progress',
         'Schedule quarterly DR training, create training materials, track completion',
         (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')],
    ]
    
    row = 5
    for idx, issue_data in enumerate(examples, start=row):
        for col_idx, value in enumerate(issue_data, start=1):
            if col_idx != 1:  # Skip Issue ID (already set)
                cell = ws.cell(row=idx, column=col_idx+1 if col_idx > 1 else col_idx, value=value)
                apply_style(cell, border=THIN_BORDER)
                if col_idx in [2, 3, 4, 7]:  # Related Test, Description, Root Cause, Remediation
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Summary metrics
    summary_row = 117
    ws.merge_cells(f'A{summary_row}:H{summary_row}')
    ws[f'A{summary_row}'] = 'ISSUE REMEDIATION SUMMARY'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    summary_row += 1
    metrics = [
        ('Total Issues:', f'=COUNTA(C5:C114)'),
        ('🔴 Critical Issues:', f'=COUNTIF(E5:E114,"🔴*")'),
        ('🟡 High Severity:', f'=COUNTIF(E5:E114,"🟡*")'),
        ('🟢 Medium Severity:', f'=COUNTIF(E5:E114,"🟢*")'),
        ('', ''),
        ('🔴 Open Issues:', f'=COUNTIF(F5:F114,"🔴*")'),
        ('⏳ In Progress:', f'=COUNTIF(F5:F114,"⏳*")'),
        (f'{CHECK} Closed Issues:', f'=COUNTIF(F5:F114,f"{CHECK}*")'),
        ('Issue Closure Rate:', f'=IF(B{summary_row}>0,B{summary_row+7}/B{summary_row},0)'),
    ]
    
    for label, formula in metrics:
        if label:
            ws[f'A{summary_row}'] = label
            apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
            ws[f'B{summary_row}'] = formula
            apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
            if 'Rate' in label:
                ws[f'B{summary_row}'].number_format = '0.0%'
        summary_row += 1
    
    # Column widths
    widths = [12, 18, 40, 35, 15, 15, 45, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'A5'
    
    return ws

# ============================================================================
# WORKSHEET: TESTING COMPLIANCE
# ============================================================================

def create_testing_compliance_sheet(wb):
    """Create Testing Compliance worksheet (annual compliance assessment)"""
    ws = wb.create_sheet(title="Testing_Compliance")
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = 'TESTING COMPLIANCE ASSESSMENT (Annual Review)'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:H2')
    ws['A2'] = 'Annual compliance assessment - are systems being tested according to requirements?'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'System Name',
        'Test Type',
        'Criticality',
        'Required Frequency',
        'Last Test Date',
        'Days Since Last Test',
        'Compliance Status',
        'Next Test Due'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    # Apply formulas and styling for 110 rows
    row = 5
    for i in range(110):
        current_row = row + i
        
        # System Name, Last Test Date (user input)
        for col in ['A', 'E']:
            if i >= 10:
                apply_style(ws[f'{col}{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)
            else:
                apply_style(ws[f'{col}{current_row}'], border=THIN_BORDER)
        ws[f'E{current_row}'].number_format = 'YYYY-MM-DD'
        
        # Test Type dropdown
        validations['test_type'].add(f'B{current_row}')
        apply_style(ws[f'B{current_row}'], border=THIN_BORDER)
        
        # Criticality dropdown
        validations['criticality'].add(f'C{current_row}')
        apply_style(ws[f'C{current_row}'], border=THIN_BORDER)
        
        # Required Frequency dropdown
        validations['test_frequency'].add(f'D{current_row}')
        apply_style(ws[f'D{current_row}'], border=THIN_BORDER)
        
        # F: Days Since Last Test (formula)
        ws[f'F{current_row}'] = f'=IF(E{current_row}="","",TODAY()-E{current_row})'
        apply_style(ws[f'F{current_row}'], border=THIN_BORDER,
                   font=Font(italic=True, color='666666'))
        
        # G: Compliance Status (formula based on frequency and days since last test)
        # Quarterly = 90 days, Semi-Annual = 180 days, Annual = 365 days
        ws[f'G{current_row}'] = f'''=IF(A{current_row}="","",IF(E{current_row}="","❓ Unknown",IF(D{current_row}="Quarterly",IF(F{current_row}<=90,f"{CHECK} Compliant",IF(F{current_row}<=105,f"{WARNING} Overdue",f"{XMARK} Non-Compliant")),IF(D{current_row}="Semi-Annual",IF(F{current_row}<=180,f"{CHECK} Compliant",IF(F{current_row}<=200,f"{WARNING} Overdue",f"{XMARK} Non-Compliant")),IF(D{current_row}="Annual",IF(F{current_row}<=365,f"{CHECK} Compliant",IF(F{current_row}<=400,f"{WARNING} Overdue",f"{XMARK} Non-Compliant")),f"{CHECK} Compliant")))))'''
        apply_style(ws[f'G{current_row}'], border=THIN_BORDER,
                   font=Font(italic=True, color='666666'))
        
        # H: Next Test Due (formula)
        ws[f'H{current_row}'] = f'''=IF(E{current_row}="","",IF(D{current_row}="Quarterly",E{current_row}+90,IF(D{current_row}="Semi-Annual",E{current_row}+180,IF(D{current_row}="Annual",E{current_row}+365,E{current_row}))))'''
        apply_style(ws[f'H{current_row}'], border=THIN_BORDER,
                   font=Font(italic=True, color='666666'))
        ws[f'H{current_row}'].number_format = 'YYYY-MM-DD'
    
    # Example compliance data (10 systems with various compliance statuses)
    examples = [
        ['E-Commerce Website', 'Backup Restore', 'Tier 1 - Critical', 'Quarterly',
         (datetime.now() - timedelta(days=70)).strftime('%Y-%m-%d')],
        ['E-Commerce Website', 'Failover Test', 'Tier 1 - Critical', 'Quarterly',
         (datetime.now() - timedelta(days=95)).strftime('%Y-%m-%d')],
        ['Payment Gateway', 'Failover Test', 'Tier 1 - Critical', 'Quarterly',
         (datetime.now() - timedelta(days=60)).strftime('%Y-%m-%d')],
        ['Email System', 'Backup Restore', 'Tier 1 - Critical', 'Quarterly',
         (datetime.now() - timedelta(days=120)).strftime('%Y-%m-%d')],
        ['CRM System', 'Backup Restore', 'Tier 1 - Critical', 'Quarterly',
         (datetime.now() - timedelta(days=80)).strftime('%Y-%m-%d')],
        ['ERP System', 'Backup Restore', 'Tier 2 - Important', 'Semi-Annual',
         (datetime.now() - timedelta(days=160)).strftime('%Y-%m-%d')],
        ['HR Information System', 'Backup Restore', 'Tier 2 - Important', 'Semi-Annual',
         (datetime.now() - timedelta(days=190)).strftime('%Y-%m-%d')],
        ['File Server', 'Backup Restore', 'Tier 3 - Standard', 'Annual',
         (datetime.now() - timedelta(days=300)).strftime('%Y-%m-%d')],
        ['Document Repository', 'Backup Restore', 'Tier 3 - Standard', 'Annual',
         (datetime.now() - timedelta(days=380)).strftime('%Y-%m-%d')],
        ['Test Database', 'Backup Restore', 'Tier 4 - Low', 'Ad-Hoc', ''],
    ]
    
    row = 5
    for idx, compliance_data in enumerate(examples, start=row):
        for col_idx, value in enumerate(compliance_data, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=value)
            apply_style(cell, border=THIN_BORDER)
    
    # Summary metrics
    summary_row = 117
    ws.merge_cells(f'A{summary_row}:H{summary_row}')
    ws[f'A{summary_row}'] = 'TESTING COMPLIANCE SUMMARY'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    summary_row += 1
    metrics = [
        ('Total Systems Tracked:', f'=COUNTA(A5:A114)'),
        (f'{CHECK} Compliant:', f'=COUNTIF(G5:G114,f"{CHECK}*")'),
        (f'{WARNING} Overdue:', f'=COUNTIF(G5:G114,f"{WARNING}*")'),
        (f'{XMARK} Non-Compliant:', f'=COUNTIF(G5:G114,f"{XMARK}*")'),
        ('❓ Unknown (Never Tested):', f'=COUNTIF(G5:G114,"❓*")'),
        ('', ''),
        ('Overall Compliance Rate:', f'=IF(B{summary_row}>0,B{summary_row+1}/B{summary_row},0)'),
        ('Tests Overdue/Non-Compliant:', f'=B{summary_row+2}+B{summary_row+3}'),
    ]
    
    for label, formula in metrics:
        if label:
            ws[f'A{summary_row}'] = label
            apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
            ws[f'B{summary_row}'] = formula
            apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
            if 'Rate' in label:
                ws[f'B{summary_row}'].number_format = '0.0%'
        summary_row += 1
    
    # Column widths
    widths = [25, 20, 18, 18, 15, 20, 20, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'A5'
    
    return ws

# ============================================================================
# WORKSHEET: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(wb):
    """Create Evidence_Register worksheet (100 rows for comprehensive audit evidence)"""
    ws = wb.create_sheet(title="Evidence_Register")
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = 'EVIDENCE REGISTER'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:H2')
    ws['A2'] = 'Document all evidence supporting this assessment (MINIMUM 5 evidence items required for audit compliance)'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'Evidence ID',
        'Evidence Type',
        'Description',
        'Related Sheet/Row',
        'Location/Path',
        'Date Collected',
        'Collected By',
        'Verification Status'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    # Evidence rows (100 rows for comprehensive documentation)
    row = 5
    for i in range(100):
        current_row = row + i
        
        # Evidence ID (auto-generated EVD-001 to EVD-100)
        ws[f'A{current_row}'] = f'EVD-{i+1:03d}'
        apply_style(ws[f'A{current_row}'], font=Font(bold=True, size=9), border=THIN_BORDER)
        
        # Evidence Type dropdown
        validations['evidence_type'].add(f'B{current_row}')
        apply_style(ws[f'B{current_row}'], border=THIN_BORDER)
        
        # Description, Related Sheet, Location, Collected By (user input)
        for col in ['C', 'D', 'E', 'G']:
            if i >= 8:  # After examples
                apply_style(ws[f'{col}{current_row}'], fill=INPUT_FILL, border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
            else:
                apply_style(ws[f'{col}{current_row}'], border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
        
        # Date Collected
        if i >= 8:
            apply_style(ws[f'F{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)
        else:
            apply_style(ws[f'F{current_row}'], border=THIN_BORDER)
        ws[f'F{current_row}'].number_format = 'YYYY-MM-DD'
        
        # Verification Status dropdown
        validations['verification_status'].add(f'H{current_row}')
        apply_style(ws[f'H{current_row}'], border=THIN_BORDER)
    
    # Example evidence (8 realistic evidence items)
    examples = [
        ['EVD-001', 'Test Report', 'E-Commerce backup restore test report - Q1 2024',
         'Test_Results_Log!A5 (RES-001)', '/evidence/testing/ecommerce_restore_Q1_2024.pdf',
         (datetime.now() - timedelta(days=5)).strftime('%Y-%m-%d'), 'Infrastructure Team', f'{CHECK} Verified'],
        ['EVD-002', 'Test Report', 'Payment Gateway failover test report with timing data',
         'Test_Results_Log!A7 (RES-003)', '/evidence/testing/payment_failover_2024-01-03.pdf',
         (datetime.now() - timedelta(days=20)).strftime('%Y-%m-%d'), 'Payments Team', f'{CHECK} Verified'],
        ['EVD-003', 'Screenshot', 'Full DR scenario test dashboard - all system status',
         'Test_Results_Log!A12 (RES-008)', '/evidence/testing/dr_scenario_dashboard_2023-10-15.png',
         (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d'), 'BC/DR Coordinator', f'{CHECK} Verified'],
        ['EVD-004', 'Video Recording', 'Tabletop exercise recording - team walkthrough',
         'Test_Results_Log!A11 (RES-007)', '/evidence/testing/tabletop_exercise_2024-01-01.mp4',
         (datetime.now() - timedelta(days=10)).strftime('%Y-%m-%d'), 'BC/DR Coordinator', f'{CHECK} Verified'],
        ['EVD-005', 'Log File', 'Email backup job failure logs - ISS-001 root cause',
         'Issue_Remediation!A5 (ISS-001)', '/evidence/logs/email_backup_failure_2023-12-10.log',
         (datetime.now() - timedelta(days=35)).strftime('%Y-%m-%d'), 'IT Operations', f'{CHECK} Verified'],
        ['EVD-006', 'Issue Ticket', 'Jira ticket ISS-006 - Payment Gateway DR scenario issue',
         'Issue_Remediation!A10 (ISS-006)', '/evidence/tickets/JIRA-BCDR-156.pdf',
         (datetime.now() - timedelta(days=85)).strftime('%Y-%m-%d'), 'BC/DR Coordinator', f'{CHECK} Verified'],
        ['EVD-007', 'Runbook', 'Updated E-Commerce restore runbook - post-test improvements',
         'Test_Results_Log!A5', '/evidence/runbooks/ecommerce_restore_runbook_v2.2.pdf',
         (datetime.now() - timedelta(days=3)).strftime('%Y-%m-%d'), 'Infrastructure Team', '⏳ Pending'],
        ['EVD-008', 'Sign-Of', 'Annual DR test sign-off from CISO',
         'Test_Results_Log!A12', '/evidence/approvals/annual_dr_test_signoff_2023.pdf',
         (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d'), 'CISO Office', f'{CHECK} Verified'],
    ]
    
    row = 5
    for idx, evidence_data in enumerate(examples, start=row):
        for col_idx, value in enumerate(evidence_data, start=1):
            if col_idx != 1:  # Skip Evidence ID (already set)
                cell = ws.cell(row=idx, column=col_idx+1 if col_idx > 1 else col_idx, value=value)
                apply_style(cell, border=THIN_BORDER)
                if col_idx in [3, 4, 5]:  # Description, Related, Location
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Evidence Summary
    summary_row = 107
    ws.merge_cells(f'A{summary_row}:H{summary_row}')
    ws[f'A{summary_row}'] = 'EVIDENCE COLLECTION SUMMARY'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    summary_row += 1
    metrics = [
        ('Total Evidence Items:', f'=COUNTA(A5:A104)'),
        (f'{CHECK} Verified Evidence:', f'=COUNTIF(H5:H104,f"{CHECK}*")'),
        ('⏳ Pending Verification:', f'=COUNTIF(H5:H104,"⏳*")'),
        (f'{XMARK} Not Verified:', f'=COUNTIF(H5:H104,f"{XMARK}*")'),
        ('', ''),
        ('Minimum Required:', '5'),
        ('Compliance Status:', f'=IF(B{summary_row}>=5,f"{CHECK} Sufficient Evidence",f"{XMARK} Insufficient Evidence")'),
    ]
    
    for label, formula in metrics:
        if label:
            ws[f'A{summary_row}'] = label
            apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
            ws[f'B{summary_row}'] = formula
            if label == 'Minimum Required:':
                apply_style(ws[f'B{summary_row}'], font=Font(bold=True, color='C00000'))
            else:
                apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
        summary_row += 1
    
    # Column widths
    widths = [12, 18, 40, 22, 35, 15, 20, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'A5'
    
    return ws

# ============================================================================
# WORKSHEET: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(wb):
    """Create Approval_Sign_Off worksheet with 3-level approval workflow"""
    ws = wb.create_sheet(title="Approval_Sign_Of")
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = 'ASSESSMENT APPROVAL & SIGN-OFF'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 30
    
    # Assessment Summary
    row = 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'ASSESSMENT SUMMARY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    summary_items = [
        ('Assessment Document:', WORKBOOK_TITLE),
        ('Assessment ID:', DOCUMENT_ID),
        ('ISO 27001:2022 Controls:', CONTROLS),
        ('Assessment Period:', '[USER INPUT - e.g., Annual 2024, Q1-Q4 2024]'),
        ('Total Tests Executed:', '=Summary!B5'),
        ('Test Success Rate:', '=Summary!B12'),
        ('Critical Issues Open:', '=Summary!B29'),
        ('Testing Compliance Rate (Tier 1):', '=Summary!B17'),
        ('Critical Issues Summary:', '[USER INPUT - list critical issues from Issue_Remediation]'),
        ('Evidence Items Collected:', '=COUNTA(Evidence_Register!A5:A104)'),
        ('Assessment Status:', '[SELECT FROM DROPDOWN]'),
    ]
    
    for label, value in summary_items:
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        cell.value = value
        
        if 'USER INPUT' in str(value) or 'SELECT' in str(value):
            apply_style(cell, fill=INPUT_FILL)
        if 'DROPDOWN' in str(value):
            validations['assessment_status'].add(cell)
        if 'Rate' in label or 'Compliance' in label:
            if '=' in str(value):
                cell.number_format = '0.0%'
        
        row += 1
    
    # Assessment Completed By (Level 1)
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'LEVEL 1: ASSESSMENT COMPLETED BY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    completion_fields = [
        ('Name:', None),
        ('Role/Title:', None),
        ('Department:', None),
        ('Email:', None),
        ('Date Completed:', 'auto_date'),
        ('Signature/Initials:', None),
    ]
    
    for field, special in completion_fields:
        ws[f'A{row}'] = field
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        apply_style(cell, fill=INPUT_FILL)
        
        if special == 'auto_date':
            cell.value = '=TODAY()'
            cell.number_format = 'YYYY-MM-DD'
            cell.font = Font(italic=True, color='808080')
        
        row += 1
    
    # Reviewed By (Level 2 - ISO)
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'LEVEL 2: REVIEWED BY (INFORMATION SECURITY OFFICER)'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    review_fields = [
        ('Name:', None),
        ('Review Date:', 'date'),
        ('Review Notes:', 3),  # Multi-line (3 rows)
        ('Recommendation:', 'recommendation_dropdown'),
    ]
    
    for field, special in review_fields:
        ws[f'A{row}'] = field
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        apply_style(cell, fill=INPUT_FILL)
        
        if special == 'recommendation_dropdown':
            validations['recommendation'].add(cell)
        elif special == 'date':
            cell.number_format = 'YYYY-MM-DD'
        elif isinstance(special, int):
            ws.merge_cells(f'B{row}:E{row+special-1}')
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            row += special - 1
        
        row += 1
    
    # Approved By (Level 3 - CISO)
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'LEVEL 3: APPROVED BY (CISO / SECURITY DIRECTOR)'
    apply_style(ws[f'A{row}'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    approval_fields = [
        ('Name:', None),
        ('Approval Date:', 'date'),
        ('Approval Decision:', 'approval_dropdown'),
        ('Conditions/Notes:', 3),  # Multi-line
    ]
    
    for field, special in approval_fields:
        ws[f'A{row}'] = field
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        apply_style(cell, fill=INPUT_FILL)
        
        if special == 'approval_dropdown':
            validations['approval_decision'].add(cell)
        elif special == 'date':
            cell.number_format = 'YYYY-MM-DD'
        elif isinstance(special, int):
            ws.merge_cells(f'B{row}:E{row+special-1}')
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            row += special - 1
        
        row += 1
    
    # Next Review Details
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'NEXT REVIEW DETAILS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    next_review_fields = [
        ('Next Annual Assessment:', 'auto_365'),
        ('Review Responsible:', None),
        ('Special Considerations:', None),
    ]
    
    for field, special in next_review_fields:
        ws[f'A{row}'] = field
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        apply_style(cell, fill=INPUT_FILL)
        
        if special == 'auto_365':
            cell.value = '=TODAY()+365'
            cell.number_format = 'YYYY-MM-DD'
            cell.font = Font(italic=True, color='808080')
        
        row += 1
    
    # Approval Workflow Notes
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'APPROVAL WORKFLOW REQUIREMENTS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    workflow_notes = [
        f'{BULLET} All 3 approval levels must be completed for final approval',
        f'{BULLET} Minimum 5 evidence items must be documented in Evidence_Register',
        f'{BULLET} Assessment status must be "Final" before CISO approval',
        f'{BULLET} All 🔴 Critical issues must have remediation plans',
        f'{BULLET} Annual re-assessment required (next assessment due in 365 days)',
        f'{BULLET} Quarterly testing required for Tier 1 systems (ongoing throughout year)',
    ]
    
    for note in workflow_notes:
        ws[f'A{row}'] = note
        ws[f'A{row}'].font = Font(name='Calibri', size=10)
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20
    
    return ws

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Generate complete BC/DR Testing Results assessment workbook"""
    
    try:
        print(f"\n{'='*70}")
        print(f"GENERATING: {WORKBOOK_TITLE}")
        print(f"{'='*70}")
        print(f"Version: {VERSION}")
        print(f"Controls: {CONTROLS}")
        print(f"Assessment ID: {DOCUMENT_ID}")
        print(f"{'='*70}\n")
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create all worksheets in order
        print("Creating worksheets...")
        create_instructions_sheet(wb)
        print("  ✅ Instructions")
        
        create_summary_sheet(wb)
        print("  ✅ Summary")
        
        create_test_schedule_sheet(wb)
        print("  ✅ Test_Schedule")
        
        create_test_results_log_sheet(wb)
        print("  ✅ Test_Results_Log")
        
        create_issue_remediation_sheet(wb)
        print("  ✅ Issue_Remediation")
        
        create_testing_compliance_sheet(wb)
        print("  ✅ Testing_Compliance")
        
        create_evidence_register(wb)
        print("  ✅ Evidence_Register")
        
        create_approval_signoff(wb)
        print("  ✅ Approval_Sign_Of")
        
        # Save workbook
        filename = f"ISMS-IMP-A.8.13.4_Testing_Results_{GENERATED_TIMESTAMP}.xlsx"
        wb.save(filename)
        
        # Summary
        print(f"\n{'='*70}")
        print("WORKBOOK GENERATED SUCCESSFULLY")
        print(f"{'='*70}")
        print(f"Filename: {filename}")
        print(f"Worksheets: {len(wb.sheetnames)}")
        print("\nWorksheet Details:")
        print("  • Instructions (comprehensive usage guide)")
        print("  • Summary (executive dashboard with all metrics)")
        print("  • Test_Schedule (110 rows: 10 examples + 100 data entry)")
        print("  • Test_Results_Log (110 rows: 10 examples + 100 results)")
        print("  • Issue_Remediation (110 rows: 10 examples + 100 issues)")
        print("  • Testing_Compliance (110 rows: 10 examples + 100 compliance checks)")
        print("  • Evidence_Register (100 rows for audit evidence, 8 examples)")
        print("  • Approval_Sign_Off (3-level workflow: Assessor → ISO → CISO)")
        print(f"\n{'='*70}")
        print(f"{CHECK} AUDIT-READY FEATURES:")
        print("  • Evidence tracking (minimum 5 items required)")
        print("  • 3-level approval workflow (Assessor → ISO → CISO)")
        print("  • Comprehensive data validations (12 dropdown types)")
        print("  • Auto-calculated compliance metrics (days since last test)")
        print("  • Test success rate tracking")
        print("  • Issue remediation workflow")
        print("  • Professional styling without Excel repair warnings")
        print(f"{'='*70}\n")
        
    except Exception as e:
        print(f"\n{'='*70}")
        print(f"{XMARK} ERROR: Failed to generate workbook")
        print(f"{'='*70}")
        print(f"Error details: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        import traceback
        print(f"\nFull traceback:")
        traceback.print_exc()
        print(f"{'='*70}\n")
        sys.exit(1)

if __name__ == "__main__":
    main()