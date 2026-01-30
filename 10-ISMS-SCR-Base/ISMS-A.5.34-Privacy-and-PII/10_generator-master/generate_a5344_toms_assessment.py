#!/usr/bin/env python3
"""
ISMS-IMP-A.5.34.4 - TOMs Assessment Generator
ISO/IEC 27001:2022 Control A.5.34: Privacy and Protection of PII
Technical and Organizational Measures (GDPR Art. 32, Swiss FADP Art. 8)
"""

import argparse, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

# Config
COLOR_HEADER = "1F4E78"
COLOR_INSTRUCTION = "D6DCE4"
COLOR_CALCULATED = "F2F2F2"
COLOR_IMPLEMENTED = "C6EFCE"
COLOR_PARTIAL = "FFEB9C"
COLOR_PLANNED = "BDD7EE"
COLOR_NOT_IMPL = "FFC7CE"
COLOR_CRITICAL = "FFC7CE"
COLOR_HIGH = "FFD966"
COLOR_MEDIUM = "FFEB9C"
COLOR_LOW = "C6EFCE"

PROTECTION_PASSWORD = "privacy2024"
FILE_PREFIX = "ISMS_A_5_34_4_TOMs_Assessment"

# 20 TOMs
TOMS = [
    ("T1", "Encryption", "Technical"),
    ("T2", "Access Control", "Technical"),
    ("T3", "Pseudonymization/Anonymization", "Technical"),
    ("T4", "Data Minimization", "Technical"),
    ("T5", "Security Monitoring & Logging", "Technical"),
    ("T6", "Network Security", "Technical"),
    ("T7", "Application Security", "Technical"),
    ("T8", "Endpoint Security", "Technical"),
    ("T9", "Backup & Recovery", "Technical"),
    ("T10", "Physical Security", "Technical"),
    ("O1", "Policies & Procedures", "Organizational"),
    ("O2", "Staff Training & Awareness", "Organizational"),
    ("O3", "Incident Response & Breach Notification", "Organizational"),
    ("O4", "Vendor Management", "Organizational"),
    ("O5", "Data Protection by Design/Default", "Organizational"),
    ("O6", "Accountability & Governance", "Organizational"),
    ("O7", "Risk Management", "Organizational"),
    ("O8", "Compliance Monitoring & Audit", "Organizational"),
    ("O9", "Documentation & Records", "Organizational"),
    ("O10", "Business Continuity", "Organizational")
]

IMPL_STATUS = ["Implemented", "Partially Implemented", "Planned", "Not Implemented"]
EFFECTIVENESS = ["Effective", "Partially Effective", "Ineffective", "Not Tested"]
RISK_LEVELS = ["Critical", "High", "Medium", "Low", "N/A"]
LIKELIHOOD = ["High", "Medium", "Low"]
IMPACT = ["Critical", "High", "Medium", "Low"]
ACTION_STATUS = ["Not Started", "In Progress", "Blocked", "Complete", "Cancelled"]

# Utilities
def create_header_row(ws, headers, row=1):
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=idx)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.protection = Protection(locked=True)

def add_dropdown(ws, col_range, values):
    dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    dv.add(col_range)
    ws.add_data_validation(dv)

def protect_sheet(ws, password=PROTECTION_PASSWORD):
    ws.protection.sheet = True
    ws.protection.password = password

# Sheet 1: Instructions
def create_sheet1(wb):
    ws = wb.active
    ws.title = "1. Instructions"
    ws['A1'] = "ISMS-IMP-A.5.34.4 - Technical and Organizational Measures (TOMs) Assessment"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    ws['A3'] = "This workbook assesses 20 TOMs protecting PII per GDPR Art. 32, Swiss FADP Art. 8, ISO 27001 A.5.34"
    ws['A3'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('A3:E5')
    ws.column_dimensions['A'].width = 100
    protect_sheet(ws)

# Sheet 2: TOM Control Inventory
def create_sheet2(wb):
    ws = wb.create_sheet("2. TOM Control Inventory")
    headers = ["TOM ID", "TOM Category", "TOM Type", "Implementation Status", "Implementation Date",
               "Description of Implementation", "Evidence Reference", "Effectiveness Rating", "Last Test Date",
               "Gaps Identified", "Risk Level", "Remediation Plan", "Remediation Owner", "Target Completion Date"]
    create_header_row(ws, headers)
    
    widths = {'A':8, 'B':35, 'C':15, 'D':20, 'E':15, 'F':80, 'G':20, 'H':20, 'I':15, 'J':60, 'K':15, 'L':60, 'M':25, 'N':15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    # Pre-fill 20 TOMs
    for idx, (tom_id, tom_cat, tom_type) in enumerate(TOMS, 2):
        ws[f'A{idx}'] = tom_id
        ws[f'A{idx}'].protection = Protection(locked=True)
        ws[f'B{idx}'] = tom_cat
        ws[f'B{idx}'].protection = Protection(locked=True)
        ws[f'C{idx}'] = tom_type
        ws[f'C{idx}'].protection = Protection(locked=True)
        for col in ['D','E','F','G','H','I','J','K','L','M','N']:
            ws[f'{col}{idx}'].protection = Protection(locked=False)
    
    # Dropdowns
    add_dropdown(ws, 'D2:D21', IMPL_STATUS)
    add_dropdown(ws, 'H2:H21', EFFECTIVENESS)
    add_dropdown(ws, 'K2:K21', RISK_LEVELS)
    
    # Conditional formatting
    ws.conditional_formatting.add('D2:D21', CellIsRule(operator='equal', formula=['"Implemented"'], 
        fill=PatternFill(start_color=COLOR_IMPLEMENTED, end_color=COLOR_IMPLEMENTED, fill_type='solid')))
    ws.conditional_formatting.add('D2:D21', CellIsRule(operator='equal', formula=['"Partially Implemented"'], 
        fill=PatternFill(start_color=COLOR_PARTIAL, end_color=COLOR_PARTIAL, fill_type='solid')))
    ws.conditional_formatting.add('D2:D21', CellIsRule(operator='equal', formula=['"Not Implemented"'], 
        fill=PatternFill(start_color=COLOR_NOT_IMPL, end_color=COLOR_NOT_IMPL, fill_type='solid')))
    
    ws.conditional_formatting.add('K2:K21', CellIsRule(operator='equal', formula=['"Critical"'], 
        fill=PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type='solid'), font=Font(bold=True)))
    ws.conditional_formatting.add('K2:K21', CellIsRule(operator='equal', formula=['"High"'], 
        fill=PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type='solid')))
    
    protect_sheet(ws)

# Sheet 3: Technical Measures Deep-Dive
def create_sheet3(wb):
    ws = wb.create_sheet("3. Technical Measures Detail")
    ws['A1'] = "TECHNICAL MEASURES DEEP-DIVE (T1-T10)"
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:C1')
    
    row = 3
    for tom_id, tom_cat, _ in [t for t in TOMS if t[0].startswith('T')]:
        ws[f'A{row}'] = f"{tom_id} - {tom_cat.upper()}"
        ws[f'A{row}'].font = Font(bold=True, size=11, color='1F4E78')
        ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
        ws.merge_cells(f'A{row}:C{row}')
        
        fields = ["Technologies Deployed:", "Configuration Details:", "Coverage (%):", "Exceptions:", "Integration Notes:"]
        for field in fields:
            row += 1
            ws[f'A{row}'] = field
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].protection = Protection(locked=False)
            ws.merge_cells(f'B{row}:C{row}')
        row += 2
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    protect_sheet(ws)

# Sheet 4: Organizational Measures Deep-Dive
def create_sheet4(wb):
    ws = wb.create_sheet("4. Organizational Measures Detail")
    ws['A1'] = "ORGANIZATIONAL MEASURES DEEP-DIVE (O1-O10)"
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:C1')
    
    row = 3
    for tom_id, tom_cat, _ in [t for t in TOMS if t[0].startswith('O')]:
        ws[f'A{row}'] = f"{tom_id} - {tom_cat.upper()}"
        ws[f'A{row}'].font = Font(bold=True, size=11, color='1F4E78')
        ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
        ws.merge_cells(f'A{row}:C{row}')
        
        fields = ["Policies/Procedures:", "Training/Communication:", "Governance Structure:", "Monitoring Method:", "Improvement Process:"]
        for field in fields:
            row += 1
            ws[f'A{row}'] = field
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].protection = Protection(locked=False)
            ws.merge_cells(f'B{row}:C{row}')
        row += 2
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    protect_sheet(ws)

# Sheet 5: Evidence Repository
def create_sheet5(wb):
    ws = wb.create_sheet("5. Evidence Repository")
    headers = ["Evidence ID", "TOM ID", "Evidence Type", "Description", "File Location", 
               "Evidence Date", "Verification Status", "Verified By", "Notes"]
    create_header_row(ws, headers)
    
    widths = {'A':15, 'B':12, 'C':35, 'D':50, 'E':60, 'F':15, 'G':20, 'H':25, 'I':50}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    evidence_types = ["Configuration Screenshot", "Policy Document", "Training Report", "Audit Report", 
                      "Test Results", "Vendor Assessment", "Incident Response Test", "Other"]
    ver_status = ["Verified", "Pending Verification", "Invalid", "Expired"]
    
    add_dropdown(ws, 'C2:C500', evidence_types)
    add_dropdown(ws, 'G2:G500', ver_status)
    
    for row in range(2, 501):
        for col in ['A','B','C','D','E','F','G','H','I']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
    
    protect_sheet(ws)

# Sheet 6: Gap Analysis
def create_sheet6(wb):
    ws = wb.create_sheet("6. Gap Analysis & Risk")
    headers = ["Gap ID", "TOM ID", "Gap Description", "Likelihood", "Impact", 
               "Overall Risk", "Risk Score", "Remediation Priority", "Residual Risk", "Acceptance Justification"]
    create_header_row(ws, headers)
    
    widths = {'A':10, 'B':12, 'C':60, 'D':15, 'E':15, 'F':15, 'G':10, 'H':20, 'I':20, 'J':50}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    add_dropdown(ws, 'D2:D200', LIKELIHOOD)
    add_dropdown(ws, 'E2:E200', IMPACT)
    
    # Risk formula
    for row in range(2, 201):
        ws[f'F{row}'] = f'=IF(AND(D{row}="High", OR(E{row}="High", E{row}="Critical")), "Critical", IF(AND(D{row}="Medium", E{row}="Critical"), "Critical", IF(OR(AND(D{row}="High", E{row}="Medium"), AND(D{row}="Medium", E{row}="High")), "High", IF(OR(AND(D{row}="Low", E{row}="High"), AND(D{row}="Medium", E{row}="Medium")), "Medium", "Low"))))'
        ws[f'F{row}'].protection = Protection(locked=True)
        
        ws[f'G{row}'] = f'=IF(F{row}="Critical", 4, IF(F{row}="High", 3, IF(F{row}="Medium", 2, 1)))'
        ws[f'G{row}'].protection = Protection(locked=True)
        
        for col in ['A','B','C','D','E','H','I','J']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
    
    ws.conditional_formatting.add('F2:F200', CellIsRule(operator='equal', formula=['"Critical"'], 
        fill=PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type='solid'), font=Font(bold=True)))
    ws.conditional_formatting.add('F2:F200', CellIsRule(operator='equal', formula=['"High"'], 
        fill=PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type='solid')))
    
    protect_sheet(ws)

# Sheet 7: Remediation Plan
def create_sheet7(wb):
    ws = wb.create_sheet("7. Remediation Action Plan")
    headers = ["Action ID", "TOM ID", "Gap Description", "Proposed Solution", "Owner", 
               "Start Date", "Target Date", "Status", "% Complete", "Completion Date"]
    create_header_row(ws, headers)
    
    widths = {'A':12, 'B':12, 'C':50, 'D':60, 'E':25, 'F':12, 'G':12, 'H':20, 'I':10, 'J':12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    add_dropdown(ws, 'H2:H200', ACTION_STATUS)
    
    for row in range(2, 201):
        for col in ['A','B','C','D','E','F','G','H','I','J']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
    
    ws.conditional_formatting.add('H2:H200', CellIsRule(operator='equal', formula=['"Complete"'], 
        fill=PatternFill(start_color=COLOR_IMPLEMENTED, end_color=COLOR_IMPLEMENTED, fill_type='solid')))
    ws.conditional_formatting.add('H2:H200', CellIsRule(operator='equal', formula=['"Blocked"'], 
        fill=PatternFill(start_color=COLOR_NOT_IMPL, end_color=COLOR_NOT_IMPL, fill_type='solid')))
    ws.conditional_formatting.add('H2:H200', FormulaRule(formula=[f'AND(H2="In Progress", G2<TODAY())'], 
        fill=PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type='solid')))
    
    protect_sheet(ws)

# Sheet 8: Dashboard
def create_sheet8(wb):
    ws = wb.create_sheet("8. Compliance Dashboard")
    ws['A1'] = "TOMS COMPLIANCE DASHBOARD"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:C1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(italic=True)
    
    metrics = [
        ("IMPLEMENTATION STATUS", [
            ("Total TOMs", "20"),
            ("Implemented", "=COUNTIF('2. TOM Control Inventory'!D:D, \"Implemented\")"),
            ("Partially Implemented", "=COUNTIF('2. TOM Control Inventory'!D:D, \"Partially Implemented\")"),
            ("Not Implemented", "=COUNTIF('2. TOM Control Inventory'!D:D, \"Not Implemented\")"),
            ("Implementation Rate", "=(B6+(B7*0.5))/B5")
        ]),
        ("CONTROL EFFECTIVENESS", [
            ("Effective Controls", "=COUNTIF('2. TOM Control Inventory'!H:H, \"Effective\")"),
            ("Effectiveness Rate", "=B12/B6")
        ]),
        ("GAP ANALYSIS", [
            ("Critical Gaps", "=COUNTIF('6. Gap Analysis & Risk'!F:F, \"Critical\")"),
            ("High Gaps", "=COUNTIF('6. Gap Analysis & Risk'!F:F, \"High\")"),
            ("Total Gaps", "=COUNTA('6. Gap Analysis & Risk'!A:A)-1")
        ]),
        ("REMEDIATION PROGRESS", [
            ("Total Actions", "=COUNTA('7. Remediation Action Plan'!A:A)-1"),
            ("Completed Actions", "=COUNTIF('7. Remediation Action Plan'!H:H, \"Complete\")"),
            ("Remediation Progress", "=B21/B20")
        ]),
        ("GDPR ART. 32 COMPLIANCE", [
            ("Compliance Score", "=(B9*0.4 + B13*0.3 + (1-B16/20)*0.2 + B22*0.1)")
        ])
    ]
    
    row = 5
    for section, items in metrics:
        ws[f'A{row}'] = section
        ws[f'A{row}'].font = Font(bold=True, size=11, color='1F4E78')
        ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        for metric_name, formula in items:
            ws[f'A{row}'] = metric_name
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = formula if formula.startswith('=') else int(formula)
            ws[f'B{row}'].protection = Protection(locked=True)
            if "Rate" in metric_name or "Progress" in metric_name or "Score" in metric_name:
                ws[f'B{row}'].number_format = '0.0%'
            row += 1
        row += 1
    
    # Compliance score formatting
    score_row = row - 2
    ws.conditional_formatting.add(f'B{score_row}', CellIsRule(operator='greaterThanOrEqual', formula=['0.9'], 
        fill=PatternFill(start_color=COLOR_IMPLEMENTED, end_color=COLOR_IMPLEMENTED, fill_type='solid'), font=Font(bold=True)))
    ws.conditional_formatting.add(f'B{score_row}', CellIsRule(operator='between', formula=['0.7', '0.9'], 
        fill=PatternFill(start_color=COLOR_PARTIAL, end_color=COLOR_PARTIAL, fill_type='solid'), font=Font(bold=True)))
    ws.conditional_formatting.add(f'B{score_row}', CellIsRule(operator='lessThan', formula=['0.7'], 
        fill=PatternFill(start_color=COLOR_NOT_IMPL, end_color=COLOR_NOT_IMPL, fill_type='solid'), font=Font(bold=True)))
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    protect_sheet(ws)

# Main
def generate_workbook(output_file):
    print("=" * 80)
    print("ISMS-IMP-A.5.34.4 - TOMs Assessment Generator")
    print("=" * 80)
    print()
    
    wb = Workbook()
    
    print("Creating Sheet 1: Instructions...")
    create_sheet1(wb)
    
    print("Creating Sheet 2: TOM Control Inventory...")
    create_sheet2(wb)
    
    print("Creating Sheet 3: Technical Measures Detail...")
    create_sheet3(wb)
    
    print("Creating Sheet 4: Organizational Measures Detail...")
    create_sheet4(wb)
    
    print("Creating Sheet 5: Evidence Repository...")
    create_sheet5(wb)
    
    print("Creating Sheet 6: Gap Analysis & Risk...")
    create_sheet6(wb)
    
    print("Creating Sheet 7: Remediation Action Plan...")
    create_sheet7(wb)
    
    print("Creating Sheet 8: Compliance Dashboard...")
    create_sheet8(wb)
    
    print()
    print(f"Saving to: {output_file}")
    wb.save(output_file)
    
    print()
    print("=" * 80)
    print("SUCCESS: Workbook generated!")
    print("=" * 80)
    print()
    print(f"Output: {os.path.abspath(output_file)}")
    print()
    return output_file

def main():
    parser = argparse.ArgumentParser(description='Generate ISMS-IMP-A.5.34.4 TOMs Assessment')
    parser.add_argument('--output', type=str, default=None, help='Output file path')
    parser.add_argument('--date', type=str, default=None, help='Date (YYYYMMDD)')
    args = parser.parse_args()
    
    date_str = args.date if args.date else datetime.now().strftime('%Y%m%d')
    output_file = args.output if args.output else f"{FILE_PREFIX}_{date_str}.xlsx"
    
    try:
        generate_workbook(output_file)
        return 0
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    exit(main())
