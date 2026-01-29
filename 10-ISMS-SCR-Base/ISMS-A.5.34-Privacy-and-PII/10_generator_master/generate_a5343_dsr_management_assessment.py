#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
===============================================================================
ISMS-IMP-A.5.34.3 - Data Subject Rights Management Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.34: Privacy and Protection of PII
Assessment Domain 3 of 7: Data Subject Rights (DSR) Management and SLA Compliance

SAMPLE SCRIPT - Requires customization for your organization's DSR procedures,
request channels, and identity verification processes.

Purpose: Generate Excel workbook for tracking and assessing data subject rights
         requests (GDPR Art. 15-22, Swiss FADP Art. 25-28) with SLA monitoring

Generated Workbook: 9 worksheets covering DSR inventory, SLA compliance, 
                    exception tracking, rights-specific analysis, and dashboard

Version: 1.0
Created: 2025-01-28
"""

import argparse, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION - CUSTOMIZE
# ============================================================================

COLOR_HEADER = "1F4E78"
COLOR_INSTRUCTION = "D6DCE4"
COLOR_INPUT = "FFFFFF"
COLOR_CALCULATED = "F2F2F2"
COLOR_SLA_MET = "C6EFCE"
COLOR_SLA_BREACHED = "FFC7CE"
COLOR_PENDING = "FFEB9C"
COLOR_WARNING = "FFD966"

PROTECTION_PASSWORD = "privacy2024"
FILE_PREFIX = "ISMS_A_5_34_3_DSR_Management_Assessment"

# CUSTOMIZE: Request channels
REQUEST_CHANNELS = ["Email", "Web Portal", "Phone", "Postal Mail", "In-Person"]

# CUSTOMIZE: The 7 data subject rights (GDPR Art. 15-22)
RIGHT_TYPES = [
    "Access (Art. 15)",
    "Rectification (Art. 16)",
    "Erasure (Art. 17)",
    "Restriction (Art. 18)",
    "Data Portability (Art. 20)",
    "Object (Art. 21)",
    "Automated Decision-Making (Art. 22)"
]

IDENTITY_VERIFICATION_METHODS = [
    "Account Login",
    "Email Confirmation",
    "ID Document",
    "Phone Verification",
    "In-Person",
    "Not Required"
]

VERIFICATION_STATUS = ["Verified", "Verification Failed", "Verification Pending", "Not Required"]

RESPONSE_METHODS = ["Email", "Secure Portal", "Postal Mail", "In-Person", "Download Link"]

REQUEST_OUTCOMES = ["Fulfilled", "Partially Fulfilled", "Rejected", "Extended", "Withdrawn"]

# CUSTOMIZE: GDPR/FADP rejection legal bases
REJECTION_REASONS = [
    "Legal Obligation (Art. 17(3)(b) - Tax, Employment Law)",
    "Legal Claims (Art. 17(3)(e) - Litigation, Defense)",
    "Public Interest (Art. 17(3)(d) - Research, Statistics)",
    "Freedom of Expression (Art. 17(3)(a))",
    "Vital Interests (Art. 17(1)(d))",
    "Manifestly Unfounded/Excessive (Art. 12(5))",
    "Overriding Legitimate Grounds (Art. 21(1))",
    "Identity Not Verified (Art. 12(6))",
    "Other (Specify in Comments)"
]

SATISFACTION_LEVELS = ["Satisfied", "Neutral", "Dissatisfied", "No Feedback"]
COMPLEXITY_LEVELS = ["Low", "Medium", "High", "Very High"]
YES_NO = ["Yes", "No"]
YES_NO_PENDING = ["Yes", "No", "Pending"]
YES_NO_NA = ["Yes", "No", "N/A"]

REQUESTER_RESPONSES = ["Accepted", "Appealed to Supervisory Authority", "Disputed", "No Response"]

EVIDENCE_TYPES = [
    "Request Email",
    "Response Email",
    "Data Export (Access/Portability)",
    "Deletion Certificate (Erasure)",
    "Identity Verification Document",
    "Legal Analysis (Rejection)",
    "Third-Party Notification (Art. 19)",
    "Appeal Documentation",
    "Extension Notification",
    "Fee Justification"
]

VERIFICATION_STATUS_EVIDENCE = ["Complete", "Incomplete", "Under Review", "Archived"]

APPROVAL_STATUS = ["Approved", "Approved with Conditions", "Requires Revision", "Rejected"]

MAX_ROWS_INVENTORY = 10000
MAX_ROWS_EXCEPTIONS = 1000
MAX_ROWS_EVIDENCE = 1000

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def create_header_row(ws, headers, start_col=1, start_row=1, header_color=COLOR_HEADER):
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_side = Side(style='thin', color='000000')
    header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for idx, header_text in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + idx)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        cell.protection = Protection(locked=True)

def set_column_widths(ws, column_widths):
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

def add_dropdown_validation(ws, col_range, values, allow_blank=False):
    dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=allow_blank)
    dv.add(col_range)
    ws.add_data_validation(dv)

def add_sla_conditional_formatting(ws, col_range):
    ws.conditional_formatting.add(col_range, CellIsRule(operator='equal', formula=['"Met"'], 
        fill=PatternFill(start_color=COLOR_SLA_MET, end_color=COLOR_SLA_MET, fill_type='solid')))
    ws.conditional_formatting.add(col_range, CellIsRule(operator='equal', formula=['"Breached"'], 
        fill=PatternFill(start_color=COLOR_SLA_BREACHED, end_color=COLOR_SLA_BREACHED, fill_type='solid'), font=Font(bold=True)))
    ws.conditional_formatting.add(col_range, CellIsRule(operator='equal', formula=['"Pending"'], 
        fill=PatternFill(start_color=COLOR_PENDING, end_color=COLOR_PENDING, fill_type='solid')))

def protect_sheet(ws, password=PROTECTION_PASSWORD):
    ws.protection.sheet = True
    ws.protection.password = password
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True

# ============================================================================
# SHEET CREATION FUNCTIONS
# ============================================================================

def create_sheet1_instructions(wb):
    ws = wb.active
    ws.title = "1. Instructions & Legend"
    
    ws['A1'] = "ISMS-IMP-A.5.34.3 - Data Subject Rights Management Assessment"
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "DSR Request Tracking and SLA Compliance Monitoring"
    ws['A2'].font = Font(name='Arial', size=11, italic=True)
    ws.merge_cells('A2:F2')
    
    row = 4
    ws[f'A{row}'] = "ASSESSMENT OVERVIEW"
    ws[f'A{row}'].font = Font(bold=True, size=12, color='1F4E78')
    
    row += 2
    overview = """This assessment tracks all data subject rights (DSR) requests to ensure compliance with:
• GDPR Articles 15-22 (Access, Rectification, Erasure, Restriction, Portability, Object, Automated Decision-Making)
• Swiss FADP Articles 25-28 (Right of access, Right to data portability, Right to rectification)
• 30-day response timeline (GDPR Art. 12(3))

CRITICAL: Non-response or delayed response = regulatory violations and fines."""
    ws[f'A{row}'] = overview
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(f'A{row}:F{row+4}')
    ws.row_dimensions[row].height = 100
    
    row += 6
    ws[f'A{row}'] = "THE 7 DATA SUBJECT RIGHTS (GDPR Art. 15-22)"
    ws[f'A{row}'].font = Font(bold=True, size=12, color='1F4E78')
    
    rights_info = [
        ("1. Access (Art. 15)", "Provide copy of personal data + processing details"),
        ("2. Rectification (Art. 16)", "Correct inaccurate personal data"),
        ("3. Erasure (Art. 17)", "Delete personal data ('right to be forgotten')"),
        ("4. Restriction (Art. 18)", "Suspend processing temporarily"),
        ("5. Portability (Art. 20)", "Receive data in machine-readable format"),
        ("6. Object (Art. 21)", "Stop specific processing (e.g., direct marketing)"),
        ("7. Automated Decision-Making (Art. 22)", "Challenge automated decisions")
    ]
    
    row += 2
    for right, description in rights_info:
        ws[f'A{row}'] = right
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'C{row}'] = description
        row += 1
    
    row += 2
    ws[f'A{row}'] = "30-DAY SLA REQUIREMENT"
    ws[f'A{row}'].font = Font(bold=True, size=12, color='1F4E78')
    
    row += 2
    sla_text = """GDPR Article 12(3): Response within 1 month (30 days) from receipt
Extension: +60 days possible if request is complex (total: 90 days)
Extension notification: Must inform data subject within 30 days with reason for delay

Direct Marketing Objection: IMMEDIATE cessation required (no 30-day grace period)"""
    ws[f'A{row}'] = sla_text
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    ws.merge_cells(f'A{row}:F{row+4}')
    
    set_column_widths(ws, {'A': 20, 'B': 15, 'C': 30, 'D': 20, 'E': 20, 'F': 20})
    
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.protection = Protection(locked=True)
    
    protect_sheet(ws)

def create_sheet2_request_inventory(wb):
    ws = wb.create_sheet("2. DSR Request Inventory")
    
    headers = [
        "Request ID", "Receipt Date", "Request Channel", "Right Type", "Requester Name",
        "Requester Contact", "Request Description", "Request Scope", "Identity Verification Method",
        "Verification Status", "Verification Date", "Response Date", "Response Deadline",
        "Days to Respond", "SLA Status", "Response Method", "Response Description", "Request Outcome",
        "Rejection Reason", "Extension Justification", "Requester Satisfaction", "Request Complexity",
        "Effort (Hours)", "Assigned To", "Evidence Reference"
    ]
    
    create_header_row(ws, headers)
    
    column_widths = {
        'A': 15, 'B': 12, 'C': 15, 'D': 20, 'E': 25, 'F': 30, 'G': 50, 'H': 40,
        'I': 20, 'J': 18, 'K': 12, 'L': 12, 'M': 12, 'N': 12, 'O': 12, 'P': 18,
        'Q': 50, 'R': 20, 'S': 40, 'T': 50, 'U': 15, 'V': 15, 'W': 10, 'X': 25, 'Y': 15
    }
    set_column_widths(ws, column_widths)
    
    # Add dropdowns
    add_dropdown_validation(ws, f'C2:C{MAX_ROWS_INVENTORY}', REQUEST_CHANNELS)
    add_dropdown_validation(ws, f'D2:D{MAX_ROWS_INVENTORY}', RIGHT_TYPES)
    add_dropdown_validation(ws, f'I2:I{MAX_ROWS_INVENTORY}', IDENTITY_VERIFICATION_METHODS)
    add_dropdown_validation(ws, f'J2:J{MAX_ROWS_INVENTORY}', VERIFICATION_STATUS)
    add_dropdown_validation(ws, f'P2:P{MAX_ROWS_INVENTORY}', RESPONSE_METHODS)
    add_dropdown_validation(ws, f'R2:R{MAX_ROWS_INVENTORY}', REQUEST_OUTCOMES)
    add_dropdown_validation(ws, f'S2:S{MAX_ROWS_INVENTORY}', REJECTION_REASONS, allow_blank=True)
    add_dropdown_validation(ws, f'U2:U{MAX_ROWS_INVENTORY}', SATISFACTION_LEVELS, allow_blank=True)
    add_dropdown_validation(ws, f'V2:V{MAX_ROWS_INVENTORY}', COMPLEXITY_LEVELS)
    
    # Add formulas
    for row in range(2, MAX_ROWS_INVENTORY + 1):
        # Response Deadline (30 days from receipt)
        ws[f'M{row}'] = f'=B{row}+30'
        ws[f'M{row}'].number_format = 'YYYY-MM-DD'
        ws[f'M{row}'].protection = Protection(locked=True)
        
        # Days to Respond
        ws[f'N{row}'] = f'=IF(L{row}="", "", L{row}-B{row})'
        ws[f'N{row}'].protection = Protection(locked=True)
        
        # SLA Status
        ws[f'O{row}'] = f'=IF(L{row}="", "Pending", IF(N{row}<=30, "Met", "Breached"))'
        ws[f'O{row}'].protection = Protection(locked=True)
    
    # Unlock input cells
    for row in range(2, MAX_ROWS_INVENTORY + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
    
    # Conditional formatting
    add_sla_conditional_formatting(ws, f'O2:O{MAX_ROWS_INVENTORY}')
    
    # Highlight overdue requests
    ws.conditional_formatting.add(f'M2:M{MAX_ROWS_INVENTORY}',
        FormulaRule(formula=[f'AND(M2<TODAY(), L2="")'], 
        fill=PatternFill(start_color=COLOR_SLA_BREACHED, end_color=COLOR_SLA_BREACHED, fill_type='solid'), font=Font(bold=True)))
    
    # Highlight verification failed
    ws.conditional_formatting.add(f'J2:J{MAX_ROWS_INVENTORY}',
        CellIsRule(operator='equal', formula=['"Verification Failed"'], 
        fill=PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type='solid')))
    
    # Highlight rejections
    ws.conditional_formatting.add(f'R2:R{MAX_ROWS_INVENTORY}',
        CellIsRule(operator='equal', formula=['"Rejected"'], 
        fill=PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type='solid')))
    
    protect_sheet(ws)

def create_sheet3_procedures(wb):
    ws = wb.create_sheet("3. Request Processing Procedures")
    
    headers = [
        "Right Type", "Standard Response Time", "Identity Verification Required",
        "Typical Fulfillment Steps", "Systems Involved", "Quality Checklist",
        "Common Exceptions", "Escalation Criteria"
    ]
    
    create_header_row(ws, headers)
    
    set_column_widths(ws, {'A': 20, 'B': 20, 'C': 15, 'D': 60, 'E': 40, 'F': 60, 'G': 50, 'H': 50})
    
    # Pre-populate rights
    for idx, right in enumerate(RIGHT_TYPES, start=2):
        ws[f'A{idx}'] = right.split('(')[0].strip()
        ws[f'A{idx}'].font = Font(bold=True)
        ws[f'A{idx}'].protection = Protection(locked=True)
    
    # Add dropdown for verification
    add_dropdown_validation(ws, 'C2:C8', ["Yes - Always", "Yes - High Risk Only", "No - Optional"])
    
    # Unlock cells for user input
    for row in range(2, 9):
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
    
    protect_sheet(ws)

def create_sheet4_sla_tracking(wb):
    ws = wb.create_sheet("4. SLA Compliance Tracking")
    
    ws['A1'] = "SLA COMPLIANCE DASHBOARD"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:D1')
    
    # Summary Metrics
    row = 4
    ws[f'A{row}'] = "SUMMARY METRICS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:B{row}')
    
    metrics = [
        ("Total Requests in Period", f"=COUNTA('2. DSR Request Inventory'!A2:A{MAX_ROWS_INVENTORY})"),
        ("Requests Met SLA", f"=COUNTIF('2. DSR Request Inventory'!O2:O{MAX_ROWS_INVENTORY}, \"Met\")"),
        ("Requests Breached SLA", f"=COUNTIF('2. DSR Request Inventory'!O2:O{MAX_ROWS_INVENTORY}, \"Breached\")"),
        ("Requests Pending", f"=COUNTIF('2. DSR Request Inventory'!O2:O{MAX_ROWS_INVENTORY}, \"Pending\")"),
        ("SLA Compliance Rate", f"=IF(B5-B8=0, 0, B6/(B5-B8))"),
        ("Average Response Time (Days)", f"=AVERAGE('2. DSR Request Inventory'!N2:N{MAX_ROWS_INVENTORY})")
    ]
    
    row += 1
    for metric_name, formula in metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = formula
        ws[f'B{row}'].protection = Protection(locked=True)
        if "Rate" in metric_name or "Percentage" in metric_name:
            ws[f'B{row}'].number_format = '0.0%'
        row += 1
    
    # Breakdown by Right Type
    row += 2
    ws[f'A{row}'] = "BREAKDOWN BY RIGHT TYPE"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    right_headers = ["Right Type", "Total", "SLA Met", "SLA Breached", "Avg Response Time"]
    for idx, header in enumerate(right_headers):
        cell = ws.cell(row=row, column=1+idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    row += 1
    for right in RIGHT_TYPES:
        right_name = right.split('(')[0].strip()
        ws[f'A{row}'] = right_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f'=COUNTIF(\'2. DSR Request Inventory\'!D2:D{MAX_ROWS_INVENTORY}, "{right}*")'
        ws[f'C{row}'] = f'=COUNTIFS(\'2. DSR Request Inventory\'!D2:D{MAX_ROWS_INVENTORY}, "{right}*", \'2. DSR Request Inventory\'!O2:O{MAX_ROWS_INVENTORY}, "Met")'
        ws[f'D{row}'] = f'=COUNTIFS(\'2. DSR Request Inventory\'!D2:D{MAX_ROWS_INVENTORY}, "{right}*", \'2. DSR Request Inventory\'!O2:O{MAX_ROWS_INVENTORY}, "Breached")'
        ws[f'E{row}'] = f'=AVERAGEIF(\'2. DSR Request Inventory\'!D2:D{MAX_ROWS_INVENTORY}, "{right}*", \'2. DSR Request Inventory\'!N2:N{MAX_ROWS_INVENTORY})'
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].protection = Protection(locked=True)
        row += 1
    
    set_column_widths(ws, {'A': 30, 'B': 15, 'C': 15, 'D': 15, 'E': 20})
    
    # Lock all cells
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.protection.locked is None:
                cell.protection = Protection(locked=True)
    
    protect_sheet(ws)

def create_sheet5_exceptions(wb):
    ws = wb.create_sheet("5. Exception & Rejection Tracking")
    
    headers = [
        "Request ID", "Right Type", "Exception Legal Basis", "Detailed Justification",
        "Data Subject Notified", "Appeal Rights Communicated", "Alternative Measures Offered",
        "DPO Review", "Legal Counsel Review", "Requester Response"
    ]
    
    create_header_row(ws, headers)
    
    set_column_widths(ws, {'A': 15, 'B': 20, 'C': 40, 'D': 60, 'E': 15, 'F': 15, 'G': 50, 'H': 15, 'I': 15, 'J': 25})
    
    add_dropdown_validation(ws, f'B2:B{MAX_ROWS_EXCEPTIONS}', RIGHT_TYPES)
    add_dropdown_validation(ws, f'C2:C{MAX_ROWS_EXCEPTIONS}', REJECTION_REASONS)
    add_dropdown_validation(ws, f'E2:E{MAX_ROWS_EXCEPTIONS}', YES_NO_PENDING)
    add_dropdown_validation(ws, f'F2:F{MAX_ROWS_EXCEPTIONS}', YES_NO)
    add_dropdown_validation(ws, f'H2:H{MAX_ROWS_EXCEPTIONS}', YES_NO_NA)
    add_dropdown_validation(ws, f'I2:I{MAX_ROWS_EXCEPTIONS}', YES_NO_NA)
    add_dropdown_validation(ws, f'J2:J{MAX_ROWS_EXCEPTIONS}', REQUESTER_RESPONSES, allow_blank=True)
    
    # Conditional formatting
    ws.conditional_formatting.add(f'F2:F{MAX_ROWS_EXCEPTIONS}',
        CellIsRule(operator='equal', formula=['"No"'], 
        fill=PatternFill(start_color=COLOR_SLA_BREACHED, end_color=COLOR_SLA_BREACHED, fill_type='solid'), font=Font(bold=True)))
    
    ws.conditional_formatting.add(f'H2:H{MAX_ROWS_EXCEPTIONS}',
        FormulaRule(formula=[f'AND(B2<>"", H2="No")'], 
        fill=PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type='solid')))
    
    for row in range(2, MAX_ROWS_EXCEPTIONS + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
    
    protect_sheet(ws)

def create_sheet6_rights_analysis(wb):
    ws = wb.create_sheet("6. Rights-Specific Analysis")
    
    ws['A1'] = "RIGHTS-SPECIFIC DEEP-DIVE ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:C1')
    
    row = 4
    for idx, right in enumerate(RIGHT_TYPES):
        right_name = right.split('(')[0].strip()
        
        ws[f'A{row}'] = right_name.upper()
        ws[f'A{row}'].font = Font(bold=True, size=12, color='1F4E78')
        ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        metrics = [
            ("Total Requests", f'=COUNTIF(\'2. DSR Request Inventory\'!D2:D{MAX_ROWS_INVENTORY}, "{right}*")'),
            ("Avg Response Time", f'=AVERAGEIF(\'2. DSR Request Inventory\'!D2:D{MAX_ROWS_INVENTORY}, "{right}*", \'2. DSR Request Inventory\'!N2:N{MAX_ROWS_INVENTORY})'),
            ("SLA Compliance Rate", f'=IFERROR(COUNTIFS(\'2. DSR Request Inventory\'!D2:D{MAX_ROWS_INVENTORY}, "{right}*", \'2. DSR Request Inventory\'!O2:O{MAX_ROWS_INVENTORY}, "Met")/COUNTIF(\'2. DSR Request Inventory\'!D2:D{MAX_ROWS_INVENTORY}, "{right}*"), 0)'),
            ("Rejection Rate", f'=IFERROR(COUNTIFS(\'2. DSR Request Inventory\'!D2:D{MAX_ROWS_INVENTORY}, "{right}*", \'2. DSR Request Inventory\'!R2:R{MAX_ROWS_INVENTORY}, "Rejected")/COUNTIF(\'2. DSR Request Inventory\'!D2:D{MAX_ROWS_INVENTORY}, "{right}*"), 0)')
        ]
        
        for metric_name, formula in metrics:
            ws[f'A{row}'] = metric_name
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = formula
            ws[f'B{row}'].protection = Protection(locked=True)
            if "Rate" in metric_name:
                ws[f'B{row}'].number_format = '0.0%'
            row += 1
        
        row += 2
    
    set_column_widths(ws, {'A': 30, 'B': 20, 'C': 20})
    
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.protection.locked is None:
                cell.protection = Protection(locked=True)
    
    protect_sheet(ws)

def create_sheet7_evidence(wb):
    ws = wb.create_sheet("7. Evidence Repository")
    
    headers = [
        "Evidence ID", "Request ID (Link)", "Evidence Type", "Description",
        "File Location", "Evidence Date", "Retention Period (Years)",
        "Verification Status", "Notes"
    ]
    
    create_header_row(ws, headers)
    
    set_column_widths(ws, {'A': 15, 'B': 15, 'C': 30, 'D': 50, 'E': 60, 'F': 12, 'G': 10, 'H': 20, 'I': 50})
    
    add_dropdown_validation(ws, f'C2:C{MAX_ROWS_EVIDENCE}', EVIDENCE_TYPES)
    add_dropdown_validation(ws, f'H2:H{MAX_ROWS_EVIDENCE}', VERIFICATION_STATUS_EVIDENCE)
    
    # Conditional formatting
    ws.conditional_formatting.add(f'H2:H{MAX_ROWS_EVIDENCE}',
        CellIsRule(operator='equal', formula=['"Incomplete"'], 
        fill=PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type='solid')))
    
    ws.conditional_formatting.add(f'F2:F{MAX_ROWS_EVIDENCE}',
        FormulaRule(formula=[f'F2+G2*365<TODAY()'], 
        fill=PatternFill(start_color=COLOR_SLA_BREACHED, end_color=COLOR_SLA_BREACHED, fill_type='solid')))
    
    for row in range(2, MAX_ROWS_EVIDENCE + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
    
    protect_sheet(ws)

def create_sheet8_dashboard(wb):
    ws = wb.create_sheet("8. Dashboard")
    
    ws['A1'] = "DSR COMPLIANCE DASHBOARD"
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(italic=True)
    ws.merge_cells('A2:D2')
    
    row = 5
    sections = [
        ("REQUEST VOLUME", [
            ("Total Requests", f"=COUNTA('2. DSR Request Inventory'!A2:A{MAX_ROWS_INVENTORY})"),
            ("Requests by Channel - Email", f"=COUNTIF('2. DSR Request Inventory'!C2:C{MAX_ROWS_INVENTORY}, \"Email\")"),
            ("Requests by Channel - Web Portal", f"=COUNTIF('2. DSR Request Inventory'!C2:C{MAX_ROWS_INVENTORY}, \"Web Portal\")")
        ]),
        ("SLA COMPLIANCE", [
            ("Overall SLA Compliance Rate", "='4. SLA Compliance Tracking'!B9"),
            ("Average Response Time (Days)", "='4. SLA Compliance Tracking'!B10"),
            ("SLA Breaches", f"=COUNTIF('2. DSR Request Inventory'!O2:O{MAX_ROWS_INVENTORY}, \"Breached\")")
        ]),
        ("REQUEST OUTCOMES", [
            ("Fulfilled", f"=COUNTIF('2. DSR Request Inventory'!R2:R{MAX_ROWS_INVENTORY}, \"Fulfilled\")"),
            ("Rejected", f"=COUNTIF('2. DSR Request Inventory'!R2:R{MAX_ROWS_INVENTORY}, \"Rejected\")"),
            ("Extended", f"=COUNTIF('2. DSR Request Inventory'!R2:R{MAX_ROWS_INVENTORY}, \"Extended\")")
        ]),
        ("EFFORT ANALYSIS", [
            ("Total Effort (Hours)", f"=SUM('2. DSR Request Inventory'!W2:W{MAX_ROWS_INVENTORY})"),
            ("Average Effort per Request", f"=AVERAGE('2. DSR Request Inventory'!W2:W{MAX_ROWS_INVENTORY})")
        ])
    ]
    
    for section_name, metrics in sections:
        ws[f'A{row}'] = section_name
        ws[f'A{row}'].font = Font(bold=True, size=12, color='1F4E78')
        ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        for metric_name, formula in metrics:
            ws[f'A{row}'] = metric_name
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = formula
            ws[f'B{row}'].protection = Protection(locked=True)
            if "Rate" in metric_name:
                ws[f'B{row}'].number_format = '0.0%'
            row += 1
        
        row += 2
    
    # Overall Compliance Score
    ws[f'A{row}'] = "OVERALL DSR COMPLIANCE SCORE"
    ws[f'A{row}'].font = Font(bold=True, size=12, color='1F4E78')
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = "Compliance Score"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'] = f"='4. SLA Compliance Tracking'!B9"
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].protection = Protection(locked=True)
    
    # Conditional formatting for compliance score
    ws.conditional_formatting.add(f'B{row}',
        CellIsRule(operator='greaterThanOrEqual', formula=['0.9'], 
        fill=PatternFill(start_color=COLOR_SLA_MET, end_color=COLOR_SLA_MET, fill_type='solid'), font=Font(bold=True)))
    ws.conditional_formatting.add(f'B{row}',
        CellIsRule(operator='between', formula=['0.7', '0.9'], 
        fill=PatternFill(start_color=COLOR_PENDING, end_color=COLOR_PENDING, fill_type='solid'), font=Font(bold=True)))
    ws.conditional_formatting.add(f'B{row}',
        CellIsRule(operator='lessThan', formula=['0.7'], 
        fill=PatternFill(start_color=COLOR_SLA_BREACHED, end_color=COLOR_SLA_BREACHED, fill_type='solid'), font=Font(bold=True, color='FFFFFF')))
    
    set_column_widths(ws, {'A': 40, 'B': 20, 'C': 20, 'D': 20})
    
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.protection.locked is None:
                cell.protection = Protection(locked=True)
    
    protect_sheet(ws)

def create_sheet9_approval(wb):
    ws = wb.create_sheet("9. Approval & Sign-Off")
    
    ws['A1'] = "DSR Assessment - Approval and Sign-Off"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    row = 4
    ws[f'A{row}'] = "ASSESSMENT COMPLETION"
    ws[f'A{row}'].font = Font(bold=True, size=12, color='1F4E78')
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:E{row}')
    
    completion_fields = [("Assessor Name:", ""), ("Role:", ""), ("Completion Date:", "")]
    row += 1
    for field, _ in completion_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].protection = Protection(locked=False)
        if "Date" in field:
            ws[f'B{row}'].number_format = 'YYYY-MM-DD'
        row += 1
    
    row += 2
    ws[f'A{row}'] = "STAKEHOLDER REVIEWS"
    ws[f'A{row}'].font = Font(bold=True, size=12, color='1F4E78')
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    stakeholder_headers = ["Stakeholder", "Name", "Date", "Approval Status", "Comments"]
    for idx, header in enumerate(stakeholder_headers):
        cell = ws.cell(row=row, column=1+idx)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    
    stakeholders = ["Data Protection Officer (DPO)", "Legal Counsel", "Customer Service Lead", "CISO / Chief Privacy Officer"]
    
    row += 1
    start_approval_row = row
    for stakeholder in stakeholders:
        ws[f'A{row}'] = stakeholder
        ws[f'A{row}'].font = Font(bold=True)
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        ws[f'C{row}'].number_format = 'YYYY-MM-DD'
        row += 1
    
    add_dropdown_validation(ws, f'D{start_approval_row}:D{row-1}', APPROVAL_STATUS)
    
    # Conditional formatting for approval status
    ws.conditional_formatting.add(f'D{start_approval_row}:D{row-1}',
        CellIsRule(operator='equal', formula=['"Approved"'], 
        fill=PatternFill(start_color=COLOR_SLA_MET, end_color=COLOR_SLA_MET, fill_type='solid')))
    ws.conditional_formatting.add(f'D{start_approval_row}:D{row-1}',
        CellIsRule(operator='equal', formula=['"Requires Revision"'], 
        fill=PatternFill(start_color=COLOR_SLA_BREACHED, end_color=COLOR_SLA_BREACHED, fill_type='solid')))
    
    row += 2
    ws[f'A{row}'] = "FINAL APPROVAL"
    ws[f'A{row}'].font = Font(bold=True, size=12, color='1F4E78')
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:E{row}')
    
    final_fields = [
        ("Final Approver:", ""), ("Title:", ""), ("Approval Date:", ""),
        ("Next Review Date:", f"=DATE(YEAR(B{row+2})+1, MONTH(B{row+2}), DAY(B{row+2}))"), ("Signature:", "")
    ]
    
    row += 1
    for field, formula_or_val in final_fields:
        ws[f'A{row}'] = field
        ws[f'A{row}'].font = Font(bold=True)
        if formula_or_val.startswith('='):
            ws[f'B{row}'] = formula_or_val
            ws[f'B{row}'].protection = Protection(locked=True)
        else:
            ws[f'B{row}'].protection = Protection(locked=False)
        if "Date" in field:
            ws[f'B{row}'].number_format = 'YYYY-MM-DD'
        row += 1
    
    set_column_widths(ws, {'A': 30, 'B': 25, 'C': 15, 'D': 25, 'E': 50})
    
    protect_sheet(ws)

# ============================================================================
# MAIN WORKBOOK GENERATION
# ============================================================================

def generate_dsr_assessment_workbook(output_file, date_str=None):
    print("=" * 80)
    print("ISMS-IMP-A.5.34.3 - DSR Management Assessment Workbook Generator")
    print("=" * 80)
    print()
    
    wb = Workbook()
    
    print("Generating Sheet 1: Instructions & Legend...")
    create_sheet1_instructions(wb)
    
    print("Generating Sheet 2: DSR Request Inventory...")
    create_sheet2_request_inventory(wb)
    
    print("Generating Sheet 3: Request Processing Procedures...")
    create_sheet3_procedures(wb)
    
    print("Generating Sheet 4: SLA Compliance Tracking...")
    create_sheet4_sla_tracking(wb)
    
    print("Generating Sheet 5: Exception & Rejection Tracking...")
    create_sheet5_exceptions(wb)
    
    print("Generating Sheet 6: Rights-Specific Analysis...")
    create_sheet6_rights_analysis(wb)
    
    print("Generating Sheet 7: Evidence Repository...")
    create_sheet7_evidence(wb)
    
    print("Generating Sheet 8: Dashboard...")
    create_sheet8_dashboard(wb)
    
    print("Generating Sheet 9: Approval & Sign-Off...")
    create_sheet9_approval(wb)
    
    print()
    print(f"Saving workbook to: {output_file}")
    wb.save(output_file)
    
    print()
    print("=" * 80)
    print("Workbook generation complete!")
    print("=" * 80)
    print()
    print("Generated workbook structure:")
    for i in range(1, 10):
        print(f"  {i}. {wb.sheetnames[i-1]}")
    print()
    print(f"Output file: {output_file}")
    print()
    
    return output_file

# ============================================================================
# CLI
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Generate ISMS-IMP-A.5.34.3 DSR Management Assessment Excel Workbook',
        epilog="""
Examples:
  python3 generate_a5343_dsr_management_assessment.py
  python3 generate_a5343_dsr_management_assessment.py --output ./dsr_assessment.xlsx
  python3 generate_a5343_dsr_management_assessment.py --date 20260128
        """
    )
    
    parser.add_argument('--output', type=str, default=None, help='Output file path')
    parser.add_argument('--date', type=str, default=None, help='Date string (YYYYMMDD)')
    
    args = parser.parse_args()
    
    date_str = args.date if args.date else datetime.now().strftime('%Y%m%d')
    output_file = args.output if args.output else f"{FILE_PREFIX}_{date_str}.xlsx"
    
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    try:
        generated_file = generate_dsr_assessment_workbook(output_file, date_str)
        print(f"SUCCESS: Workbook generated successfully")
        print(f"Location: {os.path.abspath(generated_file)}")
        return 0
    except Exception as e:
        print(f"ERROR: Failed to generate workbook: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    exit(main())
