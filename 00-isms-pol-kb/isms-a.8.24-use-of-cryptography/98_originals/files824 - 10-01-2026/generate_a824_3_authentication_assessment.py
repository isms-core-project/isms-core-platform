#!/usr/bin/env python3
"""
ISMS-IMP-A.8.24.3 - Authentication Assessment Excel Generator
ISO/IEC 27001:2022 Control A.8.24 (Use of Cryptography - Authentication)

Requirements:
    sudo apt install python3-openpyxl
    sudo apt install python3-pip

Usage:
    python3 generate_a824_3_authentication_assessment.py
    
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches markdown specification
    sheets = [
        "Instructions & Legend",
        "1. Password Security",
        "2. Multi-Factor Authentication",
        "3. Certificate-Based Auth",
        "4. Service Accounts",
        "5. SSO & Federation",
        "Summary Dashboard",
        "Evidence Register",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "exception_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "critical_alert": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        },
    }
    return styles


# ============================================================================
# SECTION 2: COLUMN DEFINITIONS (BASE + EXTENSIONS)
# ============================================================================

def get_base_auth_columns():
    """
    Return base 17 columns (A-Q) standard for ALL authentication assessment sheets.
    
    Returns:
        dict: {column_name: width}
    """
    return {
        "System/Application": 30,
        "Authentication Method": 22,
        "User Type": 18,
        "Data Classification": 18,
        "Cryptographic Algorithm": 22,
        "Hash/Encryption Status": 20,
        "Password Complexity": 20,
        "Status": 15,
        "Evidence Location": 28,
        "Gap Description": 30,
        "Remediation Needed": 14,
        "Exception ID": 14,
        "Risk ID": 14,
        "Compensating Controls": 30,
        "Responsible Person": 20,
        "Target Date": 14,
        "Budget Required": 15,
    }


def get_extended_columns(sheet_type):
    """
    Return additional columns (R-X) specific to each sheet type.
    
    Args:
        sheet_type: String identifier (password, mfa, certificate, service, sso)
    
    Returns:
        dict: {column_name: width} for columns R onwards
    """
    extensions = {
        "password": {
            "Min Password Length": 16,
            "Complexity Enforced": 16,
            "Password Expiry": 18,
            "Password History": 16,
            "Account Lockout": 16,
            "Password Strength Meter": 18,
            "Default Passwords": 18,
        },
        "mfa": {
            "MFA Factor Type": 22,
            "MFA Enrollment Rate": 18,
            "MFA Enforcement": 16,
            "Backup MFA Method": 20,
            "MFA Bypass Allowed": 18,
            "Passwordless Option": 18,
        },
        "certificate": {
            "Certificate Type": 22,
            "Issuing CA": 25,
            "Key Algorithm": 18,
            "Certificate Validity": 16,
            "CRL/OCSP Configured": 18,
            "Smart Card Required": 18,
            "PIV/CAC Compliant": 18,
        },
        "service": {
            "Service Account Type": 22,
            "Auth Method": 20,
            "Password Rotation": 18,
            "Privileged Account": 16,
            "Account Monitoring": 18,
            "Interactive Login": 18,
            "Least Privilege": 16,
        },
        "sso": {
            "SSO Protocol": 20,
            "Identity Provider": 25,
            "MFA at IdP": 16,
            "Just-In-Time Provisioning": 18,
            "Session Timeout": 18,
            "Certificate Validation": 18,
            "Token Encryption": 16,
        },
    }
    
    return extensions.get(sheet_type, {})


def get_all_columns(sheet_type):
    """Combine base + extended columns for a given sheet type."""
    base = get_base_auth_columns()
    extended = get_extended_columns(sheet_type)
    base.update(extended)
    return base


# ============================================================================
# SECTION 3: GLOBAL DATA VALIDATION DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create base dropdown validations common to all authentication sheets.
    Returns dict of DataValidation objects.
    """
    validations = {}
    
    # Authentication Method
    validations['auth_method'] = DataValidation(
        type="list",
        formula1='"Password,MFA,Certificate,Token,SSO,Federation,Biometric,API Key,Service Account,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['auth_method'])
    
    # User Type
    validations['user_type'] = DataValidation(
        type="list",
        formula1='"End User,Administrator,Service Account,External User,API Client,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['user_type'])
    
    # Data Classification
    validations['data_class'] = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Restricted,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['data_class'])
    
    # Cryptographic Algorithm
    validations['algorithm'] = DataValidation(
        type="list",
        formula1='"bcrypt,Argon2,PBKDF2,scrypt,SHA-256,SHA-512,RSA-2048+,ECDSA,None,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['algorithm'])
    
    # Hash/Encryption Status
    validations['hash_status'] = DataValidation(
        type="list",
        formula1='"Properly Hashed,Encrypted,Plaintext (Non-compliant),Salted,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['hash_status'])
    
    # Password Complexity
    validations['pwd_complexity'] = DataValidation(
        type="list",
        formula1='"Strong (≥14 chars),Adequate (12-13),Weak (<12),N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['pwd_complexity'])
    
    # Status
    validations['status'] = DataValidation(
        type="list",
        formula1='"✅ Compliant,⚠️ Partial,❌ Non-Compliant,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['status'])
    
    # Remediation Needed
    validations['remediation'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(validations['remediation'])
    
    # Budget Required
    validations['budget'] = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=False
    )
    ws.add_data_validation(validations['budget'])
    
    # Response (Yes/No/Not Applicable)
    validations['response'] = DataValidation(
        type="list",
        formula1='"Yes,No,Not Applicable"',
        allow_blank=False
    )
    ws.add_data_validation(validations['response'])
    
    # Checklist Yes/No/N/A
    validations['checklist'] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['checklist'])
    
    # Exception Yes/No
    validations['exception_yn'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(validations['exception_yn'])
    
    return validations


def create_extended_validations(ws, sheet_type):
    """Create sheet-specific dropdown validations for extended columns."""
    validations = {}
    
    if sheet_type == "password":
        validations['min_length'] = DataValidation(
            type="list",
            formula1='"≥14 chars,12-13 chars,<12 chars (weak),N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['min_length'])
        
        validations['expiry'] = DataValidation(
            type="list",
            formula1='"90 days,180 days,365 days,Never,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['expiry'])
        
        validations['history'] = DataValidation(
            type="list",
            formula1='"≥10 passwords,5-9 passwords,<5 passwords,None,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['history'])
        
        validations['default_pwd'] = DataValidation(
            type="list",
            formula1='"Changed,Not Changed (Non-compliant),N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['default_pwd'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "mfa":
        validations['mfa_factor'] = DataValidation(
            type="list",
            formula1='"TOTP (Authenticator App),Push Notification,SMS (Deprecated),Hardware Token,Biometric,Certificate,None,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['mfa_factor'])
        
        validations['enrollment'] = DataValidation(
            type="list",
            formula1='"100%,90-99%,75-89%,<75%,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['enrollment'])
        
        validations['enforcement'] = DataValidation(
            type="list",
            formula1='"Required,Optional,Not Configured,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['enforcement'])
        
        validations['bypass'] = DataValidation(
            type="list",
            formula1='"Never,With Approval,Emergency Only,Always (Non-compliant),N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['bypass'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "certificate":
        validations['cert_type'] = DataValidation(
            type="list",
            formula1='"Client Certificate,Smart Card,Machine/Device,Code Signing,Email,User,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['cert_type'])
        
        validations['cert_validity'] = DataValidation(
            type="list",
            formula1='"≤1 year,1-2 years,>2 years,Expired,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['cert_validity'])
        
        validations['key_algo'] = DataValidation(
            type="list",
            formula1='"RSA-4096,RSA-3072,RSA-2048,ECDSA P-256,ECDSA P-384,Ed25519,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['key_algo'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "service":
        validations['svc_type'] = DataValidation(
            type="list",
            formula1='"Database Service,Application Service,API Service,Scheduled Task,Integration,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['svc_type'])
        
        validations['svc_auth'] = DataValidation(
            type="list",
            formula1='"Certificate,Long Password (≥20 chars),API Key,Managed Identity,Secret,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['svc_auth'])
        
        validations['rotation'] = DataValidation(
            type="list",
            formula1='"Never (Monitored),90 days,180 days,365 days,On Change Only,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['rotation'])
        
        validations['interactive'] = DataValidation(
            type="list",
            formula1='"Disabled,Enabled (Non-compliant),N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['interactive'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "sso":
        validations['sso_protocol'] = DataValidation(
            type="list",
            formula1='"SAML 2.0,OAuth 2.0,OIDC,WS-Federation,Kerberos,LDAP,None,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['sso_protocol'])
        
        validations['session_timeout'] = DataValidation(
            type="list",
            formula1='"≤1 hour,1-8 hours,>8 hours,No timeout,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['session_timeout'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    return validations


# ============================================================================
# END OF PART 1
# ============================================================================

# ============================================================================
# SECTION 4: CHECKLIST DEFINITIONS PER AUTHENTICATION TYPE
# ============================================================================

def get_checklist_items(sheet_type):
    """
    Return checklist items for each authentication assessment type.
    
    Args:
        sheet_type: String identifier (password, mfa, certificate, service, sso)
    
    Returns:
        list: Checklist items as strings
    """
    checklists = {
        "password": [
            "Passwords hashed with approved algorithm (bcrypt, Argon2, PBKDF2, scrypt)",
            "Passwords NOT stored in plaintext anywhere",
            "Passwords NOT stored with reversible encryption",
            "Password hashes salted (unique salt per user)",
            "Minimum password length ≥12 characters",
            "Recommended password length ≥14 characters",
            "Complexity requirements enforced (uppercase, lowercase, numbers, symbols)",
            "Password expiration policy: 90 days maximum (or risk-based)",
            "Password history enforced (minimum 10 previous passwords)",
            "Account lockout after failed attempts (5-10 attempts)",
            "Lockout duration or CAPTCHA implemented",
            "Common/weak passwords blocked (breach database check)",
            "Default/vendor passwords changed before production",
            "Password strength meter provided to users",
            "Password reset process secure (no email plaintext passwords)",
            "Temporary passwords expire after first use",
            "Passwords not displayed during entry (masked)",
            "Failed login attempts logged",
            "Password policy documented and communicated",
            "Privileged accounts require stronger passwords (≥14 characters)",
        ],
        "mfa": [
            "MFA required for all administrative/privileged access",
            "MFA required for remote access (VPN, RDP, SSH)",
            "MFA required for access to Confidential/Restricted data",
            "MFA required for cloud service access (AWS, Azure, GCP, M365)",
            "MFA enrollment mandatory for all users",
            "TOTP (time-based one-time password) supported",
            "Push notification MFA supported",
            "Hardware security keys (FIDO2/WebAuthn) supported",
            "SMS-based MFA deprecated or phased out",
            "Biometric authentication available (where appropriate)",
            "Backup MFA methods configured for each user",
            "MFA bypass requests logged and approved",
            "MFA fatigue/prompt bombing protections implemented",
            "Device trust/registration for MFA",
            "MFA enrollment monitoring and alerts",
            "MFA recovery process documented",
            "Passwordless authentication option available",
            "MFA audit logs retained and reviewed",
            "Adaptive/risk-based MFA implemented",
        ],
        "certificate": [
            "Certificates issued from trusted CA (internal PKI or public CA)",
            "Certificate-based auth required for service accounts",
            "Certificate-based auth required for automated processes",
            "Client certificates used for VPN/remote access",
            "Smart cards issued for privileged users",
            "Machine/device certificates for endpoint authentication",
            "Certificate validity ≤2 years (1 year preferred)",
            "Certificate revocation checking enabled (CRL/OCSP)",
            "Certificate pinning implemented for critical services",
            "Private keys stored in HSM/TPM/smart card",
            "Private keys never exported in plaintext",
            "Certificate enrollment process documented",
            "Certificate inventory maintained",
            "Certificate expiration monitoring configured",
            "Mutual TLS (mTLS) implemented for API authentication",
            "PIV/CAC compliance (government/high-security)",
            "Certificate-based SSH authentication enabled",
            "Certificate revocation process tested",
        ],
        "service": [
            "Service account inventory maintained and up to date",
            "Certificate-based authentication used (preferred)",
            "Managed identities used for cloud services (AWS IAM, Azure MI)",
            "Long random passwords (≥20 characters) if certificate not available",
            "Password expiry disabled with compensating monitoring",
            "Service account passwords stored in secrets vault",
            "Service account passwords NEVER hardcoded in code",
            "Service account passwords NEVER in configuration files",
            "Interactive login disabled for service accounts",
            "Service accounts follow least privilege principle",
            "Service account usage logged and monitored",
            "Service account activity anomaly detection configured",
            "Privileged service accounts require additional controls",
            "Service account ownership documented",
            "Unused service accounts disabled/removed",
            "Service account naming convention followed",
            "Service account access reviewed quarterly",
            "Credential rotation tested in non-production",
            "Break-glass procedures for service account failures",
        ],
        "sso": [
            "SSO implemented for all compatible applications",
            "SAML 2.0, OAuth 2.0/OIDC, or WS-Federation protocols used",
            "Centralized Identity Provider (IdP) configured",
            "MFA enforced at IdP level",
            "SAML assertions signed and encrypted",
            "Certificate-based trust established (SAML metadata)",
            "Token lifetime/expiration configured (≤1 hour access token)",
            "Refresh token rotation implemented",
            "Just-in-time (JIT) user provisioning configured",
            "Attribute/claim mapping documented",
            "SSO session timeout configured (≤8 hours)",
            "Conditional access policies implemented",
            "Device compliance checks before SSO access",
            "SSO authentication logs centralized",
            "SSO bypass/backdoor disabled",
            "Federation trust relationships documented",
            "External federation partners vetted and approved",
            "SSO failover/redundancy configured",
            "IdP high availability implemented",
            "Privileged access management (PAM) integrated with SSO",
        ],
    }
    
    return checklists.get(sheet_type, [])


# ============================================================================
# SECTION 5: REFERENCE TABLE DEFINITIONS PER SHEET TYPE
# ============================================================================

def get_reference_tables(sheet_type):
    """
    Return reference table definitions for each sheet type.
    
    Returns:
        list: [(table_title, headers, data_rows), ...]
    """
    tables = {
        "password": [
            (
                "APPROVED PASSWORD HASHING ALGORITHMS",
                ["Algorithm", "Work Factor", "Salt Required", "Security Level", "Use Case"],
                [
                    ("Argon2id", "High (configurable)", "Yes", "Highest", "New implementations (recommended)"),
                    ("bcrypt", "Medium-High (cost factor 10-12)", "Yes", "High", "Widely supported, proven"),
                    ("PBKDF2-SHA256", "High (100,000+ iterations)", "Yes", "High", "Standards compliance (NIST)"),
                    ("scrypt", "High (configurable)", "Yes", "High", "Memory-hard function"),
                ]
            ),
            (
                "PASSWORD COMPLEXITY STANDARDS",
                ["Account Type", "Min Length", "Complexity", "Expiry", "MFA Required"],
                [
                    ("Standard User", "12 chars", "Upper, lower, number, symbol", "90 days", "Recommended"),
                    ("Administrator", "14 chars", "Upper, lower, number, symbol", "90 days", "Required"),
                    ("Service Account", "20 chars", "Random generation", "No expiry (monitor)", "Certificate preferred"),
                    ("External User", "12 chars", "Upper, lower, number, symbol", "90 days", "Required"),
                    ("API Access", "N/A", "Use API keys/tokens", "Token expiry", "Yes"),
                ]
            ),
        ],
        "mfa": [
            (
                "MFA FACTOR SECURITY COMPARISON",
                ["MFA Factor", "Security Level", "Phishing Resistant", "User Convenience", "Recommended Use"],
                [
                    ("FIDO2/WebAuthn (Hardware Key)", "Highest", "Yes", "High", "Privileged accounts, high-security"),
                    ("Certificate-based", "Highest", "Yes", "Medium", "Service accounts, automation"),
                    ("TOTP (Authenticator App)", "High", "No", "High", "General users, widely supported"),
                    ("Push Notification", "High", "Partial", "High", "General users, mobile-friendly"),
                    ("Biometric (with device trust)", "High", "Partial", "Highest", "Endpoints, mobile devices"),
                    ("SMS/Voice", "Low", "No", "Medium", "DEPRECATED - legacy only"),
                ]
            ),
            (
                "MFA IMPLEMENTATION PRIORITIES",
                ["Priority", "Account Type", "MFA Requirement", "Recommended Factor"],
                [
                    ("Critical", "Domain Administrators", "Required", "Hardware key + Backup TOTP"),
                    ("Critical", "Cloud Admins (AWS/Azure/GCP)", "Required", "Hardware key + Backup TOTP"),
                    ("Critical", "Privileged Access Management", "Required", "Hardware key or Certificate"),
                    ("High", "Standard Remote Users", "Required", "TOTP or Push"),
                    ("High", "Access to Confidential Data", "Required", "TOTP or Push"),
                    ("Medium", "Internal Users (on-premises)", "Recommended", "TOTP or Push"),
                    ("Low", "Read-only Public Data", "Optional", "TOTP"),
                ]
            ),
        ],
        "certificate": [
            (
                "CERTIFICATE AUTHENTICATION USE CASES",
                ["Use Case", "Certificate Type", "Key Algorithm", "Validity", "Storage"],
                [
                    ("Service Account Auth", "Client Certificate", "RSA-3072+", "1 year", "HSM/KMS"),
                    ("Smart Card Login", "User Certificate (PIV)", "RSA-2048+ or ECDSA P-256+", "1-3 years", "Smart Card"),
                    ("VPN Client Auth", "Client Certificate", "RSA-3072+ or ECDSA P-256+", "1 year", "TPM/Smart Card"),
                    ("Machine/Device Auth", "Device Certificate", "RSA-2048+ or ECDSA P-256+", "1-2 years", "TPM"),
                    ("Mutual TLS (mTLS)", "Client & Server Cert", "RSA-3072+ or ECDSA P-256+", "1 year", "HSM/KMS"),
                    ("Code Signing", "Code Signing Cert", "RSA-3072+", "1-3 years", "Hardware Token"),
                    ("Email Signing (S/MIME)", "Email Certificate", "RSA-2048+ or ECDSA P-256+", "1 year", "Smart Card/Keystore"),
                ]
            ),
            (
                "PKI INFRASTRUCTURE CHECKLIST",
                ["Component", "Requirement", "Implementation Status"],
                [
                    ("Root CA", "Offline, HSM-protected", ""),
                    ("Issuing CA", "Online, HSM-protected", ""),
                    ("Registration Authority", "Enrollment process automation", ""),
                    ("CRL Distribution", "HTTP/LDAP accessible", ""),
                    ("OCSP Responder", "Real-time revocation checking", ""),
                    ("Certificate Templates", "Standardized certificate profiles", ""),
                    ("Auto-enrollment", "User/machine certificate automation", ""),
                ]
            ),
        ],
        "service": [
            (
                "SERVICE ACCOUNT AUTHENTICATION METHODS",
                ["Authentication Method", "Security Level", "Use Case", "Implementation"],
                [
                    ("Managed Identity (Cloud)", "Highest", "Cloud services (AWS, Azure, GCP)", "AWS IAM Roles, Azure Managed Identity"),
                    ("Certificate-based", "Highest", "On-premises, API integration", "Client certificates, mTLS"),
                    ("Secrets Vault (HashiCorp/CyberArk)", "High", "Dynamic secret generation", "Vault dynamic credentials"),
                    ("API Key (Long, Random)", "Medium-High", "API services, integrations", "≥256-bit entropy, rotated 90 days"),
                    ("Long Password (≥20 chars)", "Medium", "Legacy systems, compatibility", "Random generation, vault storage"),
                    ("Shared Secret", "Low", "DEPRECATED", "Phase out"),
                ]
            ),
            (
                "SERVICE ACCOUNT LIFECYCLE MANAGEMENT",
                ["Lifecycle Stage", "Actions", "Responsible Role", "Frequency"],
                [
                    ("Request", "Business justification, owner assignment", "Service Owner", "As needed"),
                    ("Provisioning", "Create account, assign permissions, document", "IAM Admin", "Upon approval"),
                    ("Credential Storage", "Store in secrets vault, configure rotation", "Security Team", "At creation"),
                    ("Monitoring", "Log usage, detect anomalies, alert", "SOC/Security Team", "Continuous"),
                    ("Review", "Validate ownership, access, necessity", "IAM Admin", "Quarterly"),
                    ("Rotation", "Update credentials, test connectivity", "Service Owner", "90 days (if password)"),
                    ("Decommissioning", "Disable account, remove access, document", "IAM Admin", "Upon request"),
                ]
            ),
        ],
        "sso": [
            (
                "SSO PROTOCOL COMPARISON",
                ["Protocol", "Use Case", "Security Features", "Token Type", "Best For"],
                [
                    ("SAML 2.0", "Enterprise SSO, SaaS", "Signed/encrypted assertions, strong binding", "XML assertions", "Enterprise applications, legacy systems"),
                    ("OAuth 2.0 + OIDC", "Modern web/mobile apps, APIs", "JWT tokens, PKCE, scopes", "JSON Web Tokens (JWT)", "Cloud-native apps, mobile, APIs"),
                    ("WS-Federation", "Microsoft ecosystems", "Encrypted tokens, claims-based", "SAML tokens", "Active Directory Federation Services (ADFS)"),
                    ("Kerberos", "On-premises Windows", "Ticket-based, mutual authentication", "Kerberos tickets", "Internal Windows environments"),
                ]
            ),
            (
                "IDENTITY PROVIDER SECURITY CHECKLIST",
                ["Requirement", "Implementation Status", "Notes"],
                [
                    ("MFA enforcement at IdP", "", "Required for all users"),
                    ("Conditional access policies", "", "Risk-based authentication"),
                    ("Device compliance checking", "", "Managed devices only"),
                    ("Geo-location restrictions", "", "Block suspicious locations"),
                    ("Adaptive authentication", "", "Risk scoring, step-up auth"),
                    ("Session management", "", "Timeout, revocation"),
                    ("Audit logging", "", "All auth events logged"),
                    ("High availability", "", "99.9%+ uptime SLA"),
                ]
            ),
            (
                "TOKEN LIFETIME BEST PRACTICES",
                ["Token Type", "Recommended Lifetime", "Maximum Lifetime", "Rotation"],
                [
                    ("Access Token (OAuth)", "15 minutes", "1 hour", "Auto-expire"),
                    ("Refresh Token (OAuth)", "90 days", "1 year", "Rotate on use"),
                    ("ID Token (OIDC)", "5 minutes", "15 minutes", "Per request"),
                    ("SAML Assertion", "5 minutes", "15 minutes", "Per request"),
                    ("Session Token", "1 hour inactive", "8 hours absolute", "Extend on activity"),
                ]
            ),
        ],
    }
    
    return tables.get(sheet_type, [])


# ============================================================================
# END OF PART 2
# ============================================================================

# ============================================================================
# SECTION 6: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create the Instructions & Legend sheet matching markdown spec."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.8.24.3 – Authentication Assessment\n"
        "ISO/IEC 27001:2022 - Control A.8.24: Use of Cryptography"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.8.24.3"),
        ("Assessment Area", "Authentication Cryptographic Controls"),
        ("Related Policy", "ISMS-POL-A.8.24-S2.5"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organization", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Complete each worksheet tab (1–5) for applicable authentication methods",
        "2. Use dropdown menus for standardized entries (Status, Auth Method, Algorithm, etc.)",
        "3. Fill in yellow-highlighted cells with your information",
        "4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section",
        "5. Document password policies, MFA implementations, certificate usage, and SSO configurations",
        "6. Provide evidence location/path for each implementation entry",
        "7. Summary Dashboard auto-calculates compliance statistics per authentication area",
        "8. Maintain the Evidence Register for audit traceability",
        "9. Obtain final approval and sign-off in the Approval Sign-Off sheet",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws[f"A{row+1}"] = "STATUS LEGEND"
    ws[f"A{row+1}"].font = Font(bold=True, size=12)

    legend = [
        ("Symbol", "Status", "Description"),
        ("✅", "Compliant", "Fully meets policy requirements"),
        ("⚠️", "Partial", "Some requirements met, gaps exist"),
        ("❌", "Non-Compliant", "Does not meet policy requirements"),
        ("N/A", "Not Applicable", "Requirement does not apply"),
    ]

    row += 2
    header_row = row
    for c, h in enumerate(legend[0], start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for sym, status, desc in legend[1:]:
        ws.cell(row=row, column=1, value=sym).border = styles["border"]
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if status == "Compliant":
            s.fill = styles["status_compliant"]["fill"]
        elif status == "Partial":
            s.fill = styles["status_partial"]["fill"]
        elif status == "Non-Compliant":
            s.fill = styles["status_noncompliant"]["fill"]

        row += 1

    ws[f"A{row+2}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row+2}"].font = Font(bold=True, size=12)

    evidence_types = [
        "✓ Password policy documentation",
        "✓ Active Directory Group Policy Objects (GPO)",
        "✓ MFA enrollment reports",
        "✓ Certificate authority configurations",
        "✓ PKI infrastructure documentation",
        "✓ SSO/SAML configuration screenshots",
        "✓ Identity provider (IdP) settings",
        "✓ Service account inventory",
        "✓ Authentication logs and audit trails",
        "✓ Password hash algorithm verification",
        "✓ Account lockout policy screenshots",
        "✓ MFA bypass exception approvals",
        "✓ Federation trust relationships",
        "✓ Privilege access management (PAM) reports",
    ]
    row += 3
    for e in evidence_types:
        ws[f"A{row}"] = e
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: GENERIC ASSESSMENT SHEET ENGINE
# ============================================================================

def create_assessment_sheet(ws, styles, section_title, policy_ref, question, 
                           sheet_type):
    """
    Generic assessment sheet creator for authentication controls.
    
    Args:
        ws: worksheet object
        styles: style dictionary
        section_title: e.g., "1. PASSWORD SECURITY - PASSWORD AUTHENTICATION CONTROLS"
        policy_ref: policy requirement text
        question: assessment question
        sheet_type: key for columns/checklist/tables (password, mfa, certificate, service, sso)
    """
    columns = get_all_columns(sheet_type)
    checklist_items = get_checklist_items(sheet_type)
    reference_tables = get_reference_tables(sheet_type)
    
    total_cols = len(columns)
    last_col_letter = get_column_letter(total_cols)
    
    # ---------- HEADER ----------
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"{section_title}\nPolicy Requirement: {policy_ref}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # ---------- ASSESSMENT QUESTION ----------
    ws.merge_cells(f"A3:{last_col_letter}3")
    ws["A3"] = question
    ws["A3"].font = Font(bold=True, size=11)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # ---------- RESPONSE DROPDOWN ----------
    ws["A4"] = "Response:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    # Create validations
    base_validations = create_base_validations(ws)
    extended_validations = create_extended_validations(ws, sheet_type)
    base_validations['response'].add(ws["B4"])

    # ---------- COLUMN HEADERS ----------
    header_row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ---------- EXAMPLE ROW ----------
    example_row = 7
    example_values = _get_example_row(sheet_type, total_cols)
    for col_idx, value in enumerate(example_values, start=1):
        cell = ws.cell(row=example_row, column=col_idx, value=value)
        cell.font = Font(italic=True, color="808080")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---------- DATA ENTRY ROWS (8-20) ----------
    start_row = 8
    end_row = 20
    
    for r in range(start_row, end_row + 1):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # Apply base validations to standard columns (B-Q, indices 2-17)
    for r in range(start_row, end_row + 1):
        base_validations['auth_method'].add(ws.cell(row=r, column=2))  # B: Authentication Method
        base_validations['user_type'].add(ws.cell(row=r, column=3))    # C: User Type
        base_validations['data_class'].add(ws.cell(row=r, column=4))   # D: Data Classification
        base_validations['algorithm'].add(ws.cell(row=r, column=5))    # E: Cryptographic Algorithm
        base_validations['hash_status'].add(ws.cell(row=r, column=6))  # F: Hash/Encryption Status
        base_validations['pwd_complexity'].add(ws.cell(row=r, column=7)) # G: Password Complexity
        base_validations['status'].add(ws.cell(row=r, column=8))       # H: Status
        base_validations['remediation'].add(ws.cell(row=r, column=11)) # K: Remediation Needed
        base_validations['budget'].add(ws.cell(row=r, column=17))      # Q: Budget Required

    # Apply extended validations for columns R onwards (18+)
    _apply_extended_validations(ws, sheet_type, extended_validations, start_row, end_row)

    ws.freeze_panes = "A7"

    next_row = end_row + 2

    # ---------- COMPLIANCE CHECKLIST ----------
    ws[f"A{next_row}"] = f"{section_title.split('-')[0].strip().upper()} CHECKLIST"
    ws[f"A{next_row}"].font = Font(bold=True, size=11)
    next_row += 1

    ws[f"A{next_row}"] = "☐"
    ws[f"B{next_row}"] = "Requirement"
    ws[f"C{next_row}"] = "Status"
    for col in ["A", "B", "C"]:
        ws[f"{col}{next_row}"].font = Font(bold=True)
    next_row += 1

    checklist_start = next_row
    for item in checklist_items:
        ws[f"A{next_row}"] = "☐"
        ws[f"B{next_row}"] = item
        ws[f"C{next_row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{next_row}"].border = styles["border"]

        base_validations['checklist'].add(ws[f"C{next_row}"])
        next_row += 1

    # Checklist score
    ws[f"A{next_row}"] = "Checklist Score:"
    ws[f"A{next_row}"].font = Font(bold=True)
    ws[f"B{next_row}"] = (
        f'=COUNTIF(C{checklist_start}:C{next_row-1},"Yes")/'
        f'COUNTA(C{checklist_start}:C{next_row-1})*100&"%"'
    )
    ws[f"B{next_row}"].font = Font(bold=True, color="0000FF")
    next_row += 2

    # ---------- REFERENCE TABLES ----------
    for table_title, headers, data_rows in reference_tables:
        ws[f"A{next_row}"] = table_title
        ws[f"A{next_row}"].font = Font(bold=True, size=11)
        next_row += 1
        
        # Table headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=header)
            cell.font = styles["column_header"]["font"]
            cell.fill = styles["column_header"]["fill"]
            cell.alignment = styles["column_header"]["alignment"]
            cell.border = styles["column_header"]["border"]
        next_row += 1
        
        # Table data rows
        for data_row in data_rows:
            for col_idx, value in enumerate(data_row, start=1):
                cell = ws.cell(row=next_row, column=col_idx, value=value)
                cell.border = styles["border"]
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # For PKI Infrastructure table (has status dropdown in column 3)
                if "PKI INFRASTRUCTURE" in table_title and col_idx == 3:
                    cell.fill = styles["input_cell"]["fill"]
                    # Create dropdown for implementation status
                    dv_impl = DataValidation(
                        type="list",
                        formula1='"Implemented,Planned,Not Implemented"',
                        allow_blank=False
                    )
                    ws.add_data_validation(dv_impl)
                    dv_impl.add(cell)
            next_row += 1
        
        next_row += 1  # Space between tables

    # ---------- EXCEPTION/DEVIATION BLOCK ----------
    ws.merge_cells(f"A{next_row}:{last_col_letter}{next_row}")
    ws[f"A{next_row}"] = "EXCEPTION / DEVIATION DOCUMENTATION (Complete if Status = Partial or Non-Compliant)"
    ws[f"A{next_row}"].font = styles["exception_header"]["font"]
    ws[f"A{next_row}"].fill = styles["exception_header"]["fill"]
    ws[f"A{next_row}"].alignment = styles["exception_header"]["alignment"]

    next_row += 1
    exception_fields = [
        ("Formal exception request submitted:", "Yes,No"),
        ("Exception ID:", ""),
        ("Risk acceptance documented:", "Yes,No"),
        ("Risk ID:", ""),
        ("Compensating Controls (summary):", ""),
        ("☐ Enhanced monitoring and alerting", ""),
        ("☐ Network segmentation / restricted access", ""),
        ("☐ Additional authentication factor", ""),
        ("☐ Manual approval process", ""),
        ("☐ Increased audit frequency", ""),
        ("☐ Time-limited exception (expiry date required)", ""),
        ("☐ Other (describe):", ""),
    ]

    for label, options in exception_fields:
        ws[f"A{next_row}"] = label
        ws[f"A{next_row}"].font = Font(bold=True)
        ws.merge_cells(f"B{next_row}:D{next_row}")
        ws[f"B{next_row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{next_row}"].border = styles["border"]
        ws[f"B{next_row}"].alignment = styles["input_cell"]["alignment"]

        if options:
            base_validations['exception_yn'].add(ws[f"B{next_row}"])

        next_row += 1

    # ---------- NOTES ----------
    next_row += 1
    ws.merge_cells(f"A{next_row}:{last_col_letter}{next_row}")
    ws[f"A{next_row}"] = "ADDITIONAL NOTES / COMMENTS"
    ws[f"A{next_row}"].font = Font(bold=True, size=11)

    next_row += 1
    ws.merge_cells(f"A{next_row}:{last_col_letter}{next_row+8}")
    ws[f"A{next_row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{next_row}"].border = styles["border"]
    ws[f"A{next_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Column widths for checklist/tables
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 20

    return (start_row, end_row, 8)  # Return status column index (H = 8)


def _get_example_row(sheet_type, total_cols):
    """Generate example row values based on sheet type."""
    examples = {
        "password": [
            "Example: Production HR System - Workday", "Password", "End User", "Confidential",
            "bcrypt", "Properly Hashed", "Strong (≥14 chars)", "✅ Compliant",
            "/evidence/auth/workday-password-policy.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "≥14 chars", "Yes", "90 days", "≥10 passwords", "Yes", "Yes", "Changed"
        ],
        "mfa": [
            "Example: AWS Production Account", "MFA", "Administrator", "Restricted",
            "TOTP", "Encrypted", "N/A", "✅ Compliant",
            "/evidence/auth/aws-mfa-config.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "TOTP (Authenticator App)", "100%", "Required", "Yes", "Never", "Available"
        ],
        "certificate": [
            "Example: VPN Gateway - Cisco ASA", "Certificate", "End User", "Confidential",
            "RSA-3072", "Encrypted", "N/A", "✅ Compliant",
            "/evidence/auth/vpn-cert-config.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "Client Certificate", "Internal CA", "RSA-3072", "≤1 year", "Yes", "No", "No"
        ],
        "service": [
            "Example: Database Service Account - SQL_SVC", "Service Account", "Service Account", "Restricted",
            "Certificate", "Encrypted", "N/A", "✅ Compliant",
            "/evidence/auth/sqlsvc-cert.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "Database Service", "Certificate", "Never (Monitored)", "Yes", "Yes", "Disabled", "Yes"
        ],
        "sso": [
            "Example: Office 365 SSO - Azure AD", "SSO", "End User", "Confidential",
            "SHA-256", "Encrypted", "N/A", "✅ Compliant",
            "/evidence/auth/o365-sso-config.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "SAML 2.0", "Azure AD", "Yes", "Enabled", "≤1 hour", "Yes", "Yes"
        ],
    }
    
    base_example = examples.get(sheet_type, ["Example"] * total_cols)
    # Pad or trim to match total_cols
    if len(base_example) < total_cols:
        base_example.extend(["N/A"] * (total_cols - len(base_example)))
    return base_example[:total_cols]


def _apply_extended_validations(ws, sheet_type, validations, start_row, end_row):
    """Apply extended column validations based on sheet type."""
    if sheet_type == "password":
        for r in range(start_row, end_row + 1):
            validations['min_length'].add(ws.cell(row=r, column=18))  # R
            validations['yn_na'].add(ws.cell(row=r, column=19))       # S: Complexity Enforced
            validations['expiry'].add(ws.cell(row=r, column=20))      # T
            validations['history'].add(ws.cell(row=r, column=21))     # U
            validations['yn_na'].add(ws.cell(row=r, column=22))       # V: Account Lockout
            validations['yn_na'].add(ws.cell(row=r, column=23))       # W: Password Strength Meter
            validations['default_pwd'].add(ws.cell(row=r, column=24)) # X
    
    elif sheet_type == "mfa":
        for r in range(start_row, end_row + 1):
            validations['mfa_factor'].add(ws.cell(row=r, column=18))  # R
            validations['enrollment'].add(ws.cell(row=r, column=19))  # S
            validations['enforcement'].add(ws.cell(row=r, column=20)) # T
            validations['yn_na'].add(ws.cell(row=r, column=21))       # U: Backup MFA Method
            validations['bypass'].add(ws.cell(row=r, column=22))      # V
            validations['yn_na'].add(ws.cell(row=r, column=23))       # W: Passwordless Option
    
    elif sheet_type == "certificate":
        for r in range(start_row, end_row + 1):
            validations['cert_type'].add(ws.cell(row=r, column=18))   # R
            # S: Issuing CA - free text
            validations['key_algo'].add(ws.cell(row=r, column=20))    # T
            validations['cert_validity'].add(ws.cell(row=r, column=21)) # U
            validations['yn_na'].add(ws.cell(row=r, column=22))       # V: CRL/OCSP
            validations['yn_na'].add(ws.cell(row=r, column=23))       # W: Smart Card
            validations['yn_na'].add(ws.cell(row=r, column=24))       # X: PIV/CAC
    
    elif sheet_type == "service":
        for r in range(start_row, end_row + 1):
            validations['svc_type'].add(ws.cell(row=r, column=18))    # R
            validations['svc_auth'].add(ws.cell(row=r, column=19))    # S
            validations['rotation'].add(ws.cell(row=r, column=20))    # T
            validations['yn_na'].add(ws.cell(row=r, column=21))       # U: Privileged Account
            validations['yn_na'].add(ws.cell(row=r, column=22))       # V: Account Monitoring
            validations['interactive'].add(ws.cell(row=r, column=23)) # W
            validations['yn_na'].add(ws.cell(row=r, column=24))       # X: Least Privilege
    
    elif sheet_type == "sso":
        for r in range(start_row, end_row + 1):
            validations['sso_protocol'].add(ws.cell(row=r, column=18)) # R
            # S: Identity Provider - free text
            validations['yn_na'].add(ws.cell(row=r, column=20))        # T: MFA at IdP
            validations['yn_na'].add(ws.cell(row=r, column=21))        # U: JIT Provisioning
            validations['session_timeout'].add(ws.cell(row=r, column=22)) # V
            validations['yn_na'].add(ws.cell(row=r, column=23))        # W: Certificate Validation
            validations['yn_na'].add(ws.cell(row=r, column=24))        # X: Token Encryption


# ============================================================================
# END OF PART 3
# ============================================================================

# ============================================================================
# SECTION 8: INDIVIDUAL ASSESSMENT SHEET DEFINITIONS
# ============================================================================

def create_1_password_security(ws, styles):
    """1. Password Security - matches markdown spec exactly."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="1. PASSWORD SECURITY - PASSWORD AUTHENTICATION CONTROLS",
        policy_ref="Passwords MUST be hashed using approved algorithms (bcrypt, Argon2, PBKDF2, scrypt). Minimum 12 characters, complexity requirements enforced (Policy Section 2.5.1)",
        question="Does your organization use password-based authentication for any systems or applications?",
        sheet_type="password",
    )


def create_2_multi_factor_authentication(ws, styles):
    """2. Multi-Factor Authentication - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="2. MULTI-FACTOR AUTHENTICATION (MFA) - STRONG AUTHENTICATION",
        policy_ref="MFA REQUIRED for all administrative access, remote access, and access to Confidential/Restricted data. Approved factors: TOTP, Push notification, Hardware token, Biometric (Policy Section 2.5.2)",
        question="Does your organization require multi-factor authentication for any systems or user types?",
        sheet_type="mfa",
    )


def create_3_certificate_based_auth(ws, styles):
    """3. Certificate-Based Authentication - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="3. CERTIFICATE-BASED AUTHENTICATION - PKI & DIGITAL CERTIFICATES",
        policy_ref="Certificate-based authentication REQUIRED for service accounts, automated processes, and high-security environments. Certificates issued from trusted CA with proper key management (Policy Section 2.5.3)",
        question="Does your organization use digital certificates for authentication (client certificates, smart cards, machine authentication)?",
        sheet_type="certificate",
    )


def create_4_service_accounts(ws, styles):
    """4. Service Accounts - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="4. SERVICE ACCOUNTS - NON-HUMAN AUTHENTICATION",
        policy_ref="Service accounts MUST use certificate-based authentication or long random passwords (≥20 characters). Password expiry disabled with monitoring. MFA or certificate preferred (Policy Section 2.5.4)",
        question="Does your organization use service accounts for application-to-application or automated process authentication?",
        sheet_type="service",
    )


def create_5_sso_federation(ws, styles):
    """5. SSO & Federation - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="5. SSO & FEDERATION - CENTRALIZED AUTHENTICATION",
        policy_ref="SSO/Federation REQUIRED using SAML 2.0, OAuth 2.0/OIDC, or WS-Federation. Identity Provider (IdP) must enforce strong authentication and MFA (Policy Section 2.5.5)",
        question="Does your organization use Single Sign-On (SSO) or federated authentication for applications and services?",
        sheet_type="sso",
    )


# ============================================================================
# SECTION 9: SUMMARY DASHBOARD (COMPLEX - 12 ANALYSIS SECTIONS)
# ============================================================================

def create_summary_dashboard(ws, styles):
    """Create comprehensive summary dashboard with 12 analysis sections."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "AUTHENTICATION ASSESSMENT - COMPLIANCE SUMMARY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # ---------- 1. COMPLIANCE SUMMARY TABLE ----------
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    areas = [
        ("1. Password Security", "1. Password Security"),
        ("2. Multi-Factor Authentication", "2. Multi-Factor Authentication"),
        ("3. Certificate-Based Auth", "3. Certificate-Based Auth"),
        ("4. Service Accounts", "4. Service Accounts"),
        ("5. SSO & Federation", "5. SSO & Federation"),
    ]

    row += 1
    start_data_row = row
    for label, sheet in areas:
        ws.cell(row=row, column=1, value=label)
        
        status_range = f"'{sheet}'!H8:H20"

        ws.cell(row=row, column=2, value=f'=COUNTA({status_range})')
        ws.cell(row=row, column=3, value=f'=COUNTIF({status_range},"✅*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({status_range},"⚠️*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({status_range},"❌*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({status_range},"N/A")')
        ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    # Total row
    total_row = row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{row-1})")
        cell.font = Font(bold=True)

    total_pct = ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
    total_pct.font = Font(bold=True, color="0000FF", size=12)

    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 15

    # ---------- 2. AUTHENTICATION METHOD DISTRIBUTION ----------
    row += 3
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "AUTHENTICATION METHOD DISTRIBUTION"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    method_headers = ["Authentication Method", "Count", "Percentage", "Security Rating"]
    for col_idx, header in enumerate(method_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    methods = [
        ("Password Only", "LOW RISK ⚠️"),
        ("Password + MFA", "Good ✅"),
        ("Certificate-based", "Excellent ✅"),
        ("SSO + MFA", "Excellent ✅"),
        ("Biometric + MFA", "Excellent ✅"),
        ("Legacy/Weak", "CRITICAL RISK ❌"),
    ]

    row += 1
    method_start = row
    for method, rating in methods:
        ws.cell(row=row, column=1, value=method)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual count
        ws.cell(row=row, column=3, value=f'=IF(SUM(B{method_start}:B{row})=0,"0%",ROUND(B{row}/SUM(B${method_start}:B${row})*100,1)&"%")')
        ws.cell(row=row, column=4, value=rating)
        row += 1

    # ---------- 3. MFA ADOPTION METRICS ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "MFA ADOPTION METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    mfa_headers = ["User Type", "Total Users", "MFA Enrolled", "Enrollment %", "MFA Enforcement"]
    for col_idx, header in enumerate(mfa_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    user_types = ["Administrators", "Remote Users", "Internal Users", "External Users", "Service Accounts"]
    row += 1
    for utype in user_types:
        ws.cell(row=row, column=1, value=utype)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual
        ws.cell(row=row, column=3).fill = styles["input_cell"]["fill"]  # Manual
        ws.cell(row=row, column=4, value=f'=IF(B{row}>0,C{row}/B{row}*100&"%","N/A")')
        ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]  # Manual: Required/Optional
        row += 1

    # ---------- 4. PASSWORD SECURITY METRICS ----------
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "PASSWORD SECURITY METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    pwd_headers = ["Password Metric", "Count", "Percentage", "Compliance Status"]
    for col_idx, header in enumerate(pwd_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    pwd_metrics = [
        ("Hashed with Approved Algorithm", "✅ Compliant"),
        ("Plaintext Storage", "❌ CRITICAL"),
        ("Weak Hashing (MD5/SHA-1)", "❌ Non-Compliant"),
        ("Min Length ≥14 chars", "✅ Recommended"),
        ("Min Length 12-13 chars", "⚠️ Adequate"),
        ("Min Length <12 chars", "❌ Weak"),
        ("Account Lockout Enabled", "✅ Required"),
        ("Password Expiry Configured", "✅ Required"),
    ]

    row += 1
    pwd_start = row
    for metric, status in pwd_metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual/formula
        ws.cell(row=row, column=3, value=f'=IF(SUM(B{pwd_start}:B{row})=0,"0%",ROUND(B{row}/SUM(B${pwd_start}:B${row})*100,1)&"%")')
        ws.cell(row=row, column=4, value=status)
        row += 1

    # ---------- 5. SERVICE ACCOUNT SECURITY ----------
    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "SERVICE ACCOUNT SECURITY"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    svc_headers = ["Service Account Metric", "Count", "Status"]
    for col_idx, header in enumerate(svc_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    svc_metrics = [
        ("Total Service Accounts", ""),
        ("Certificate-based Auth", "✅"),
        ("Long Password (≥20 chars)", "✅"),
        ("Weak Password (<20 chars)", "❌"),
        ("Interactive Login Disabled", "✅"),
        ("Stored in Secrets Vault", "✅"),
        ("Privileged Accounts", "⚠️"),
    ]

    row += 1
    for metric, status in svc_metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual count
        ws.cell(row=row, column=3, value=status)
        row += 1

    # ---------- 6. SSO COVERAGE ANALYSIS ----------
    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "SSO COVERAGE ANALYSIS"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    sso_headers = ["SSO Metric", "Count", "Status"]
    for col_idx, header in enumerate(sso_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    sso_metrics = [
        ("Applications with SSO", ""),
        ("SAML 2.0 Implementations", ""),
        ("OAuth 2.0/OIDC Implementations", ""),
        ("MFA Required at IdP", "✅"),
        ("Legacy Protocols (LDAP/Kerberos)", "⚠️"),
        ("JIT Provisioning Enabled", "✅"),
    ]

    row += 1
    for metric, status in sso_metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual
        ws.cell(row=row, column=3, value=status)
        row += 1

    # ---------- 7. AUTHENTICATION SECURITY SCORE ----------
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "OVERALL AUTHENTICATION SECURITY SCORE"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Component"
    ws[f"B{row}"] = "Weight"
    ws[f"C{row}"] = "Score"
    ws[f"D{row}"] = "Weighted Score"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].font = Font(bold=True)

    components = [
        ("Password Security (20%)", "0.20"),
        ("MFA Coverage (30%)", "0.30"),
        ("Certificate Auth (15%)", "0.15"),
        ("Service Accounts (15%)", "0.15"),
        ("SSO & Federation (20%)", "0.20"),
    ]

    row += 1
    score_start = row
    for idx, (comp, weight) in enumerate(components, start=1):
        ws.cell(row=row, column=1, value=comp)
        ws.cell(row=row, column=2, value=weight)
        # Reference checklist scores from each sheet
        sheet_idx = idx  # 1-5
        ws.cell(row=row, column=3).fill = styles["input_cell"]["fill"]  # Manual or formula
        ws.cell(row=row, column=4, value=f"=C{row}*B{row}")
        row += 1

    ws[f"A{row}"] = "TOTAL WEIGHTED SCORE:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"D{row}"] = f"=SUM(D{score_start}:D{row-1})*100&\"%\""
    ws[f"D{row}"].font = Font(bold=True, size=12, color="0000FF")

    row += 1
    ws[f"A{row}"] = "Security Posture Rating:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Dropdown: Excellent/Good/Adequate/Needs Improvement

    # ---------- 8. CRITICAL SECURITY GAPS ----------
    row += 3
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "CRITICAL GAPS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    gap_headers = ["Priority", "Assessment Area", "Gap Description", "Risk Level", "Responsible Person", "Target Date", "Status"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=False)
    ws.add_data_validation(dv_priority)
    
    dv_risk = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=False)
    ws.add_data_validation(dv_risk)

    dv_gap_status = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Closed"', allow_blank=False)
    ws.add_data_validation(dv_gap_status)

    for i in range(8):
        row += 1
        ws.cell(row=row, column=1).fill = styles["input_cell"]["fill"]
        dv_priority.add(ws.cell(row=row, column=1))
        
        for col in range(2, 6):
            c = ws.cell(row=row, column=col)
            c.fill = styles["input_cell"]["fill"]
            c.border = styles["border"]
            c.alignment = styles["input_cell"]["alignment"]
        
        dv_risk.add(ws.cell(row=row, column=4))
        
        ws.cell(row=row, column=6).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=7).fill = styles["input_cell"]["fill"]
        dv_gap_status.add(ws.cell(row=row, column=7))

    ws.freeze_panes = "A4"


# ============================================================================
# END OF PART 4
# ============================================================================

# ============================================================================
# SECTION 10: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create evidence register for audit traceability."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID",
        "Assessment Area",
        "Evidence Type",
        "Description",
        "Location/Path",
        "Date Collected",
        "Collected By",
        "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Validation dropdowns
    dv_type = DataValidation(
        type="list",
        formula1='"Password policy,GPO screenshot,MFA enrollment report,Certificate config,PKI documentation,SAML configuration,IdP settings,Service account inventory,Authentication logs,Audit trail,Policy document,Security assessment,Penetration test,Compliance scan,Other"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_type)

    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_ver)

    # Data rows (100 rows as per spec)
    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EV-{r-4:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 11: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create approval and sign-off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", "ISMS-IMP-A.8.24.3 - Authentication Assessment"),
        ("Assessment Period", ""),
        ("Overall Compliance Rate", "='Summary Dashboard'!G8"),  # Total row from dashboard
        ("Assessment Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)

        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Status dropdown
    status_cell_row = row - 1
    dv_status = DataValidation(
        type="list", 
        formula1='"Draft,Final,Requires remediation,Re-assessment required"', 
        allow_blank=False
    )
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{status_cell_row}"])

    # ---------- ASSESSMENT COMPLETED BY ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    completion_fields = ["Name", "Role/Title", "Department", "Email", "Date"]
    row += 1
    for field in completion_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # ---------- REVIEWED BY (INFORMATION SECURITY OFFICER) ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (INFORMATION SECURITY OFFICER)"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    review_fields = ["Name", "Date", "Review Notes"]
    row += 1
    for field in review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # ---------- APPROVED BY (CISO) ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    approval_fields = ["Name", "Date", "Approval Decision", "Conditions/Notes"]
    row += 1
    for field in approval_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Approval Decision dropdown (3rd field from approval section start)
    decision_cell_row = row - 2
    dv_dec = DataValidation(
        type="list", 
        formula1='"Approved,Approved with conditions,Rejected"', 
        allow_blank=False
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{decision_cell_row}"])

    # ---------- NEXT REVIEW DETAILS ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    next_review_fields = ["Next Review Date", "Review Responsible", "Special Considerations"]
    row += 1
    for field in next_review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    print("=" * 78)
    print("ISMS-IMP-A.8.24.3 - Authentication Assessment Generator")
    print("ISO/IEC 27001:2022 Control A.8.24: Use of Cryptography (Authentication)")
    print("=" * 78)

    wb = create_workbook()
    styles = setup_styles()

    print("\n[1/9] Creating Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)

    print("[2/9] Creating 1. Password Security sheet...")
    create_1_password_security(wb["1. Password Security"], styles)

    print("[3/9] Creating 2. Multi-Factor Authentication sheet...")
    create_2_multi_factor_authentication(wb["2. Multi-Factor Authentication"], styles)

    print("[4/9] Creating 3. Certificate-Based Auth sheet...")
    create_3_certificate_based_auth(wb["3. Certificate-Based Auth"], styles)

    print("[5/9] Creating 4. Service Accounts sheet...")
    create_4_service_accounts(wb["4. Service Accounts"], styles)

    print("[6/9] Creating 5. SSO & Federation sheet...")
    create_5_sso_federation(wb["5. SSO & Federation"], styles)

    print("[7/9] Creating Summary Dashboard sheet...")
    create_summary_dashboard(wb["Summary Dashboard"], styles)

    print("[8/9] Creating Evidence Register sheet...")
    create_evidence_register(wb["Evidence Register"], styles)

    print("[9/9] Creating Approval Sign-Off sheet...")
    create_approval_signoff(wb["Approval Sign-Off"], styles)

    filename = f"ISMS-IMP-A.8.24.3_Authentication_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    print(f"\n✅ SUCCESS: {filename}")
    print("\nWorkbook Structure:")
    print("  • Instructions & Legend - Document info and guidance")
    print("  • 5 Assessment Sheets:")
    print("    1. Password Security (20 checklist items + 2 reference tables)")
    print("    2. Multi-Factor Authentication (19 checklist items + 2 reference tables)")
    print("    3. Certificate-Based Auth (18 checklist items + 2 reference tables)")
    print("    4. Service Accounts (19 checklist items + 2 reference tables)")
    print("    5. SSO & Federation (20 checklist items + 3 reference tables)")
    print("  • Summary Dashboard - 8 comprehensive analysis sections:")
    print("    - Compliance summary by authentication type")
    print("    - Authentication method distribution")
    print("    - MFA adoption metrics by user type")
    print("    - Password security metrics")
    print("    - Service account security analysis")
    print("    - SSO coverage analysis")
    print("    - Overall authentication security score (weighted)")
    print("    - Critical security gaps (8 entry rows)")
    print("  • Evidence Register - 100 evidence entry rows")
    print("  • Approval Sign-Off - Complete workflow with 3 sign-off levels")
    
    print("\nColumn Structure:")
    print("  • Base columns (A-Q): 17 standard columns across all sheets")
    print("  • Extended columns (R-X):")
    print("    - Password Security: +7 columns (min length, complexity, expiry, etc.)")
    print("    - MFA: +6 columns (factor type, enrollment, enforcement, etc.)")
    print("    - Certificate: +7 columns (cert type, CA, key algo, validity, etc.)")
    print("    - Service Accounts: +7 columns (account type, auth method, rotation, etc.)")
    print("    - SSO & Federation: +7 columns (protocol, IdP, MFA, JIT, session, etc.)")
    
    print("\nNext steps:")
    print("  1) Complete document information in Instructions & Legend")
    print("  2) Fill yellow cells in each assessment sheet (1–5)")
    print("  3) Check compliance checklists per authentication type")
    print("  4) Review reference tables (algorithms, use cases, best practices)")
    print("  5) Document exceptions/deviations as needed")
    print("  6) Maintain Evidence Register entries")
    print("  7) Review Summary Dashboard:")
    print("     - Overall compliance by authentication area")
    print("     - Authentication method distribution and risk ratings")
    print("     - MFA adoption rates by user type")
    print("     - Password security compliance metrics")
    print("     - Service account security posture")
    print("     - SSO coverage and protocol usage")
    print("     - Weighted authentication security score")
    print("     - Critical gaps requiring immediate remediation")
    print("  8) Complete multi-level approval sign-off workflow")
    print("\n" + "=" * 78)


if __name__ == "__main__":
    main()


# ============================================================================
# END OF SCRIPT - COMPLETE AUTHENTICATION ASSESSMENT GENERATOR
# ============================================================================