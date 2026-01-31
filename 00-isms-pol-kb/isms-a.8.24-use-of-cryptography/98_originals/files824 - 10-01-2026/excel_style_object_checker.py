#!/usr/bin/env python3
"""
Excel Style Object Uniqueness Checker - Detects shared object issues
Checks for shared Border, Font, and Fill objects - common causes of Excel "repair" warnings

GENERIC VERSION - Works for all Excel workbooks

Usage:
    python3 excel_style_object_checker.py ISMS-IMP-A.8.24.1_Data_Transmission_20251231.xlsx
    python3 excel_style_object_checker.py --full ISMS-IMP-A.8.24.5_Compliance_Summary_Dashboard_20251231.xlsx
    
"""

import sys
from openpyxl import load_workbook


# ============================================================================
# WORKBOOK TYPE DETECTION
# ============================================================================

def detect_workbook_type(filename):
    """Detect which A.8.24 workbook type this is based on filename."""
    filename_lower = filename.lower()
    
    if 'a.8.24.1' in filename_lower or 'data_transmission' in filename_lower:
        return 'A.8.24.1', 'Data Transmission Assessment', 100
    elif 'a.8.24.2' in filename_lower or 'data_storage' in filename_lower:
        return 'A.8.24.2', 'Data Storage Assessment', 100
    elif 'a.8.24.3' in filename_lower or 'authentication' in filename_lower:
        return 'A.8.24.3', 'Authentication Assessment', 100
    elif 'a.8.24.4' in filename_lower or 'key_management' in filename_lower:
        return 'A.8.24.4', 'Key Management Assessment', 100
    elif 'a.8.24.5' in filename_lower or 'compliance_summary' in filename_lower or 'dashboard' in filename_lower:
        return 'A.8.24.5', 'Compliance Summary Dashboard', 200
    else:
        return 'Unknown', 'Generic Excel Workbook', 100


# ============================================================================
# STYLE OBJECT CHECKING
# ============================================================================

def check_style_objects(filename, full_scan=False):
    """
    Check if style objects (borders, fonts, fills) are being reused.
    
    Args:
        filename: Excel file to check
        full_scan: If True, scan all rows; if False, scan intelligently based on workbook type
    """
    
    workbook_id, workbook_name, default_rows = detect_workbook_type(filename)
    
    print("=" * 80)
    print(f"EXCEL STYLE OBJECT UNIQUENESS CHECK: {filename}")
    print(f"Detected Type: {workbook_id} - {workbook_name}")
    print("=" * 80)
    
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"\n✓ Workbook loaded successfully")
        print(f"  Sheets found: {len(wb.sheetnames)}")
    except Exception as e:
        print(f"\n✗ CRITICAL: Cannot load workbook: {e}")
        return
    
    # Determine scan depth
    if full_scan:
        max_rows = None
        print(f"  Scan mode: FULL (all rows)")
    else:
        max_rows = default_rows
        print(f"  Scan mode: SMART (first {max_rows} rows per sheet)")
    
    print("\nNote: Use --full flag to scan all rows (slower but more thorough)")
    
    # Statistics collectors
    border_stats = {
        'total_cells': 0,
        'unique_objects': set(),
        'sheets': {}
    }
    
    font_stats = {
        'total_cells': 0,
        'unique_objects': set(),
        'sheets': {}
    }
    
    fill_stats = {
        'total_cells': 0,
        'unique_objects': set(),
        'sheets': {}
    }
    
    print("\nScanning for style object references...")
    print("(This may take a moment for large workbooks)\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Initialize sheet stats
        border_stats['sheets'][sheet_name] = {'cells': 0, 'unique': set()}
        font_stats['sheets'][sheet_name] = {'cells': 0, 'unique': set()}
        fill_stats['sheets'][sheet_name] = {'cells': 0, 'unique': set()}
        
        # Determine row range
        if max_rows:
            row_range = ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row))
        else:
            row_range = ws.iter_rows()
        
        for row in row_range:
            for cell in row:
                # Check borders
                if hasattr(cell, 'border') and cell.border is not None:
                    border_stats['total_cells'] += 1
                    border_stats['sheets'][sheet_name]['cells'] += 1
                    border_id = id(cell.border)
                    border_stats['unique_objects'].add(border_id)
                    border_stats['sheets'][sheet_name]['unique'].add(border_id)
                
                # Check fonts
                if hasattr(cell, 'font') and cell.font is not None:
                    font_stats['total_cells'] += 1
                    font_stats['sheets'][sheet_name]['cells'] += 1
                    font_id = id(cell.font)
                    font_stats['unique_objects'].add(font_id)
                    font_stats['sheets'][sheet_name]['unique'].add(font_id)
                
                # Check fills
                if hasattr(cell, 'fill') and cell.fill is not None:
                    fill_stats['total_cells'] += 1
                    fill_stats['sheets'][sheet_name]['cells'] += 1
                    fill_id = id(cell.fill)
                    fill_stats['unique_objects'].add(fill_id)
                    fill_stats['sheets'][sheet_name]['unique'].add(fill_id)
    
    # ========================================================================
    # BORDER ANALYSIS
    # ========================================================================
    print("=" * 80)
    print("BORDER OBJECT ANALYSIS")
    print("=" * 80)
    
    critical_sheets = []
    warning_sheets = []
    
    for sheet_name, stats in border_stats['sheets'].items():
        if stats['cells'] > 0:
            reuse_ratio = stats['cells'] / len(stats['unique']) if stats['unique'] else 0
            print(f"\n  {sheet_name}:")
            print(f"    Cells with borders: {stats['cells']}")
            print(f"    Unique border objects: {len(stats['unique'])}")
            print(f"    Reuse ratio: {reuse_ratio:.1f}x", end='')
            
            if reuse_ratio > 10:
                print(f"  ❌ CRITICAL")
                critical_sheets.append(sheet_name)
            elif reuse_ratio > 5:
                print(f"  ⚠️  WARNING")
                warning_sheets.append(sheet_name)
            else:
                print(f"  ✓ OK")
    
    # Overall border summary
    if border_stats['total_cells'] > 0:
        overall_border_reuse = border_stats['total_cells'] / len(border_stats['unique_objects'])
        print(f"\n  Overall Border Statistics:")
        print(f"    Total cells with borders: {border_stats['total_cells']}")
        print(f"    Unique border objects: {len(border_stats['unique_objects'])}")
        print(f"    Overall reuse ratio: {overall_border_reuse:.1f}x")
    
    # ========================================================================
    # FONT ANALYSIS
    # ========================================================================
    print("\n" + "=" * 80)
    print("FONT OBJECT ANALYSIS")
    print("=" * 80)
    
    font_critical = []
    font_warning = []
    
    for sheet_name, stats in font_stats['sheets'].items():
        if stats['cells'] > 0:
            reuse_ratio = stats['cells'] / len(stats['unique']) if stats['unique'] else 0
            
            # Only show sheets with concerning reuse
            if reuse_ratio > 5:
                print(f"\n  {sheet_name}:")
                print(f"    Cells with fonts: {stats['cells']}")
                print(f"    Unique font objects: {len(stats['unique'])}")
                print(f"    Reuse ratio: {reuse_ratio:.1f}x", end='')
                
                if reuse_ratio > 20:
                    print(f"  ❌ CRITICAL")
                    font_critical.append(sheet_name)
                else:
                    print(f"  ⚠️  WARNING")
                    font_warning.append(sheet_name)
    
    if not font_critical and not font_warning:
        print("\n  ✓ All font object reuse within acceptable limits")
    
    if font_stats['total_cells'] > 0:
        overall_font_reuse = font_stats['total_cells'] / len(font_stats['unique_objects'])
        print(f"\n  Overall Font Statistics:")
        print(f"    Total cells with fonts: {font_stats['total_cells']}")
        print(f"    Unique font objects: {len(font_stats['unique_objects'])}")
        print(f"    Overall reuse ratio: {overall_font_reuse:.1f}x")
    
    # ========================================================================
    # FILL ANALYSIS
    # ========================================================================
    print("\n" + "=" * 80)
    print("FILL OBJECT ANALYSIS")
    print("=" * 80)
    
    fill_critical = []
    fill_warning = []
    
    for sheet_name, stats in fill_stats['sheets'].items():
        if stats['cells'] > 0:
            reuse_ratio = stats['cells'] / len(stats['unique']) if stats['unique'] else 0
            
            # Only show sheets with concerning reuse
            if reuse_ratio > 5:
                print(f"\n  {sheet_name}:")
                print(f"    Cells with fills: {stats['cells']}")
                print(f"    Unique fill objects: {len(stats['unique'])}")
                print(f"    Reuse ratio: {reuse_ratio:.1f}x", end='')
                
                if reuse_ratio > 20:
                    print(f"  ❌ CRITICAL")
                    fill_critical.append(sheet_name)
                else:
                    print(f"  ⚠️  WARNING")
                    fill_warning.append(sheet_name)
    
    if not fill_critical and not fill_warning:
        print("\n  ✓ All fill object reuse within acceptable limits")
    
    if fill_stats['total_cells'] > 0:
        overall_fill_reuse = fill_stats['total_cells'] / len(fill_stats['unique_objects'])
        print(f"\n  Overall Fill Statistics:")
        print(f"    Total cells with fills: {fill_stats['total_cells']}")
        print(f"    Unique fill objects: {len(fill_stats['unique_objects'])}")
        print(f"    Overall reuse ratio: {overall_fill_reuse:.1f}x")
    
    # ========================================================================
    # FINAL DIAGNOSIS
    # ========================================================================
    print("\n" + "=" * 80)
    print("DIAGNOSIS & RECOMMENDATIONS")
    print("=" * 80)
    
    total_issues = len(critical_sheets) + len(font_critical) + len(fill_critical)
    total_warnings = len(warning_sheets) + len(font_warning) + len(fill_warning)
    
    if total_issues == 0 and total_warnings == 0:
        print("\n✅ EXCELLENT: NO SHARED STYLE OBJECTS DETECTED")
        print("\nAll style objects (borders, fonts, fills) are unique per cell.")
        print("This is not the cause of Excel repair warnings.")
        print("\nIf Excel still shows repair warnings, investigate:")
        print("  • Formula errors (run excel_sanity_check.py)")
        print("  • Data validation conflicts")
        print("  • Merged cell issues")
        print("  • Excel version compatibility")
    
    elif total_issues > 0:
        print(f"\n❌ CRITICAL: {total_issues} SHARED OBJECT ISSUE(S) DETECTED")
        
        if critical_sheets:
            print(f"\n🔴 BORDER ISSUES ({len(critical_sheets)} sheets):")
            for sheet in critical_sheets:
                print(f"   • {sheet}")
        
        if font_critical:
            print(f"\n🔴 FONT ISSUES ({len(font_critical)} sheets):")
            for sheet in font_critical:
                print(f"   • {sheet}")
        
        if fill_critical:
            print(f"\n🔴 FILL ISSUES ({len(fill_critical)} sheets):")
            for sheet in fill_critical:
                print(f"   • {sheet}")
        
        print("\n" + "=" * 80)
        print("📋 FIX REQUIRED - UPDATE YOUR GENERATOR SCRIPT:")
        print("=" * 80)
        
        print("\nYour script is creating style objects ONCE and reusing them.")
        print("This is the #1 cause of Excel repair warnings with openpyxl.")
        
        print("\n🔧 SOLUTION:")
        print("   Replace shared style objects with factory functions")
        
        print("\n   # ❌ OLD (WRONG - creates shared objects):")
        print("   border_thin = Border(left=Side(style='thin'), ...)")
        print("   font_bold = Font(bold=True)")
        print("   fill_yellow = PatternFill(start_color='FFFFCC', ...)")
        print("   ")
        print("   cell.border = border_thin  # <- Shared reference!")
        print("   cell.font = font_bold      # <- Shared reference!")
        print("   cell.fill = fill_yellow    # <- Shared reference!")
        
        print("\n   # ✅ NEW (CORRECT - creates unique objects):")
        print("   def create_thin_border():")
        print("       thin = Side(style='thin')")
        print("       return Border(left=thin, right=thin, top=thin, bottom=thin)")
        print("   ")
        print("   def create_bold_font():")
        print("       return Font(bold=True)")
        print("   ")
        print("   def create_yellow_fill():")
        print("       return PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')")
        print("   ")
        print("   cell.border = create_thin_border()  # <- Unique object!")
        print("   cell.font = create_bold_font()      # <- Unique object!")
        print("   cell.fill = create_yellow_fill()    # <- Unique object!")
        
        print("\n   💡 KEY POINT: Call the function for EACH cell")
        print("      Each function call creates a NEW, UNIQUE object")
        
        print("\n📝 STEPS TO FIX:")
        print("   1. Update your generator script with factory functions")
        print("   2. Replace all 'cell.border = border_thin' with 'cell.border = create_thin_border()'")
        print("   3. Replace all 'cell.font = font_bold' with 'cell.font = create_bold_font()'")
        print("   4. Replace all 'cell.fill = fill_yellow' with 'cell.fill = create_yellow_fill()'")
        print("   5. Regenerate the workbook")
        print("   6. Run this checker again to verify")
    
    elif total_warnings > 0:
        print(f"\n⚠️  WARNING: {total_warnings} MODERATE STYLE OBJECT REUSE DETECTED")
        print("\nSome style objects are being shared across multiple cells.")
        print("This may contribute to Excel repair warnings.")
        print("\nConsider updating your generator script to use unique objects per cell.")
        print("See fix instructions above for critical issues.")
    
    print("\n" + "=" * 80)


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 style_object_checker.py [--full] <filename.xlsx>")
        print("\nExamples:")
        print("  python3 style_object_checker.py ISMS-IMP-A.8.24.1_Data_Transmission_20251231.xlsx")
        print("  python3 style_object_checker.py --full ISMS-IMP-A.8.24.5_Compliance_Summary_Dashboard_20251231.xlsx")
        print("\nOptions:")
        print("  --full    Scan all rows (slower but more thorough)")
        print("            Default: Smart scan (first 100-200 rows based on workbook type)")
        sys.exit(1)
    
    # Parse arguments
    full_scan = False
    filename = None
    
    for arg in sys.argv[1:]:
        if arg == '--full':
            full_scan = True
        else:
            filename = arg
    
    if not filename:
        print("Error: No filename provided")
        sys.exit(1)
    
    check_style_objects(filename, full_scan)


if __name__ == "__main__":
    main()