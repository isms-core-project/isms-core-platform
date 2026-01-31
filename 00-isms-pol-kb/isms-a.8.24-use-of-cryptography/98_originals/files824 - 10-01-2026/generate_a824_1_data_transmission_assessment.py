#!/usr/bin/env python3
"""
ISMS-IMP-A.8.24.1 - Data Transmission Assessment Excel Generator
ISO/IEC 27001:2022 Control A.8.24 (Use of Cryptography)

Requirements:
    sudo apt install python3-openpyxl
    sudo apt install python3-pip

Usage:
    python3 generate_a824_1_data_transmission_assessment.py
    
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches markdown subsections
    sheets = [
        "Instructions & Legend",
        "1.1 External HTTPS-TLS",
        "1.2 Internal HTTPS-TLS",
        "2.1 Email Encryption",
        "2.2 Digital Signatures",
        "3.1 File Transfer Protocols",
        "4.1 VPN",
        "4.2 SSH",
        "4.3 RDP",
        "5.1 API Security",
        "6.1 Database Connections",
        "6.2 Wireless Networks",
        "7.1 Cloud Transmission",
        "Summary Dashboard",
        "Evidence Register",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# ============================================================================
# SECTION 2: COLUMN DEFINITIONS PER SUBSECTION
# ============================================================================

def get_column_definitions(section_key):
    """
    Return column definitions (name: width) for each subsection.
    Matches the markdown specification exactly.
    """
    column_defs = {
        "1.1_external_https": {
            "Service Description": 35,
            "Current TLS Version": 18,
            "Certificate Source (CA)": 22,
            "Certificate Validity": 16,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "1.2_internal_https": {
            "Service Description": 35,
            "Data Classification": 18,
            "Current TLS Version": 18,
            "Certificate Type": 20,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "2.1_email_encryption": {
            "Email System": 25,
            "Encryption Solution": 22,
            "Encryption Method": 20,
            "User Adoption Rate": 16,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "2.2_digital_signatures": {
            "Use Case": 30,
            "Signature Method": 20,
            "Certificate Source": 22,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "3.1_file_transfer": {
            "Transfer Method/System": 30,
            "Protocol Used": 18,
            "Authentication Method": 22,
            "Data Classification": 18,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "4.1_vpn": {
            "VPN Solution": 25,
            "Protocol": 18,
            "Encryption Algorithm": 20,
            "MFA Enabled": 14,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "4.2_ssh": {
            "System/Service": 30,
            "SSH Version": 14,
            "Authentication Method": 22,
            "Key Algorithm": 18,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "4.3_rdp": {
            "System/Environment": 30,
            "RDP Access Method": 22,
            "TLS Encryption": 16,
            "NLA Enabled": 14,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "5.1_api": {
            "API Name/Service": 30,
            "Authentication Method": 22,
            "TLS Version": 14,
            "API Key Management": 22,
            "Token Expiry": 16,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "6.1_database": {
            "Database System": 25,
            "Connection Encryption": 22,
            "Certificate Validation": 20,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "6.2_wireless": {
            "Network SSID": 25,
            "Encryption Standard": 20,
            "Authentication Method": 22,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
        "7.1_cloud": {
            "Cloud Provider/Service": 30,
            "Connection Method": 22,
            "Encryption": 18,
            "Status": 15,
            "Evidence Location": 30,
            "Gap Description": 30,
            "Remediation Needed": 16,
        },
    }
    
    return column_defs.get(section_key, {})


# ============================================================================
# END OF PART 1
# ============================================================================

# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create the Instructions & Legend sheet matching markdown spec."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.8.24.1 – Data Transmission Assessment\n"
        "ISO/IEC 27001:2022 - Control A.8.24: Use of Cryptography"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.8.24.1"),
        ("Assessment Area", "Data Transmission Cryptographic Controls"),
        ("Related Policy", "ISMS-POL-A.8.24-S2.2"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organization", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Complete each worksheet tab for applicable systems/services.",
        "2. Use dropdown menus for standardized entries (Status, Remediation, etc.).",
        "3. Fill in yellow-highlighted cells with your information.",
        "4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section.",
        "5. Check all applicable items in the Compliance Checklist for each section.",
        "6. Provide evidence location/path for each implementation entry.",
        "7. Summary Dashboard auto-calculates compliance statistics.",
        "8. Maintain the Evidence Register for audit traceability.",
        "9. Obtain final approval and sign-off in the Approval Sign-Off sheet.",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws[f"A{row+1}"] = "STATUS LEGEND"
    ws[f"A{row+1}"].font = Font(bold=True, size=12)

    legend = [
        ("Symbol", "Status", "Description"),
        ("✅", "Compliant", "Fully meets policy requirements"),
        ("⚠️", "Partial", "Some requirements met, gaps exist"),
        ("❌", "Non-Compliant", "Does not meet policy requirements"),
        ("N/A", "Not Applicable", "Requirement does not apply"),
    ]

    row += 2
    header_row = row
    for c, h in enumerate(legend[0], start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for sym, status, desc in legend[1:]:
        ws.cell(row=row, column=1, value=sym).border = styles["border"]
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if status == "Compliant":
            s.fill = styles["status_compliant"]["fill"]
        elif status == "Partial":
            s.fill = styles["status_partial"]["fill"]
        elif status == "Non-Compliant":
            s.fill = styles["status_noncompliant"]["fill"]

        row += 1

    ws[f"A{row+2}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row+2}"].font = Font(bold=True, size=12)

    evidence_types = [
        "✓ Configuration files or screenshots",
        "✓ Network scan results (e.g., SSL Labs, nmap, testssl.sh)",
        "✓ System documentation / architecture diagrams",
        "✓ Vendor specifications / encryption statements",
        "✓ Certificate inventory / key management records",
        "✓ Audit logs / SIEM evidence",
        "✓ Compliance reports / exception approvals",
    ]
    row += 3
    for e in evidence_types:
        ws[f"A{row}"] = e
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 4: GENERIC ASSESSMENT SHEET ENGINE
# ============================================================================

def create_assessment_sheet(ws, styles, section_title, policy_ref, question, 
                           section_key, checklist_items, additional_details):
    """
    Generic assessment sheet creator.
    
    Args:
        ws: worksheet object
        styles: style dictionary
        section_title: e.g., "1.1 HTTPS/TLS - External Web Services"
        policy_ref: policy requirement text
        question: assessment question
        section_key: key for column definitions (e.g., "1.1_external_https")
        checklist_items: list of compliance checklist items
        additional_details: list of (label, dropdown_options) tuples for Additional Details section
    """
    columns = get_column_definitions(section_key)
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{section_title}\nPolicy Requirement: {policy_ref}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # ---------- ASSESSMENT QUESTION ----------
    ws.merge_cells("A3:H3")
    ws["A3"] = question
    ws["A3"].font = Font(bold=True, size=11)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # ---------- RESPONSE DROPDOWN ----------
    ws["A4"] = "Response:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    dv_response = DataValidation(type="list", formula1='"Yes,No,Not Applicable"', allow_blank=False)
    ws.add_data_validation(dv_response)
    dv_response.add(ws["B4"])

    # ---------- COLUMN HEADERS ----------
    header_row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ---------- EXAMPLE ROW ----------
    example_row = 7
    example_values = ["Example entry (see markdown spec)" if i == 0 else "" 
                     for i in range(len(columns))]
    for col_idx, value in enumerate(example_values, start=1):
        cell = ws.cell(row=example_row, column=col_idx, value=value)
        cell.font = Font(italic=True, color="808080")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---------- DATA ENTRY ROWS (8-20) ----------
    start_row = 8
    end_row = 20
    status_col_idx = list(columns.keys()).index("Status") + 1
    remediation_col_idx = list(columns.keys()).index("Remediation Needed") + 1
    
    for r in range(start_row, end_row + 1):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # Status dropdown
    dv_status = DataValidation(type="list", 
                               formula1='"✅ Compliant,⚠️ Partial,❌ Non-Compliant,N/A"', 
                               allow_blank=False)
    ws.add_data_validation(dv_status)
    for r in range(start_row, end_row + 1):
        dv_status.add(ws.cell(row=r, column=status_col_idx))

    # Remediation Needed dropdown
    dv_rem = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv_rem)
    for r in range(start_row, end_row + 1):
        dv_rem.add(ws.cell(row=r, column=remediation_col_idx))

    ws.freeze_panes = "A7"

    next_row = end_row + 2

    # ---------- ADDITIONAL DETAILS ----------
    if additional_details:
        ws[f"A{next_row}"] = "Additional Details"
        ws[f"A{next_row}"].font = Font(bold=True, size=11)
        next_row += 1
        
        for label, options in additional_details:
            ws[f"A{next_row}"] = label
            ws[f"A{next_row}"].font = Font(bold=True)
            ws.merge_cells(f"B{next_row}:D{next_row}")
            ws[f"B{next_row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{next_row}"].border = styles["border"]
            ws[f"B{next_row}"].alignment = styles["input_cell"]["alignment"]
            
            if options:
                dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=False)
                ws.add_data_validation(dv)
                dv.add(ws[f"B{next_row}"])
            
            next_row += 1
        
        next_row += 1

    # ---------- COMPLIANCE CHECKLIST ----------
    if checklist_items:
        ws[f"A{next_row}"] = "COMPLIANCE CHECKLIST"
        ws[f"A{next_row}"].font = Font(bold=True, size=11)
        next_row += 1

        ws[f"A{next_row}"] = "☐"
        ws[f"B{next_row}"] = "Requirement"
        ws[f"C{next_row}"] = "Status"
        for col in ["A", "B", "C"]:
            ws[f"{col}{next_row}"].font = Font(bold=True)
        next_row += 1

        checklist_start = next_row
        for item in checklist_items:
            ws[f"A{next_row}"] = "☐"
            ws[f"B{next_row}"] = item
            ws[f"C{next_row}"].fill = styles["input_cell"]["fill"]
            ws[f"C{next_row}"].border = styles["border"]

            dv_chk = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=False)
            ws.add_data_validation(dv_chk)
            dv_chk.add(ws[f"C{next_row}"])
            next_row += 1

        # Checklist score
        ws[f"A{next_row}"] = "Checklist Score:"
        ws[f"A{next_row}"].font = Font(bold=True)
        ws[f"B{next_row}"] = (
            f'=COUNTIF(C{checklist_start}:C{next_row-1},"Yes")/'
            f'COUNTA(C{checklist_start}:C{next_row-1})*100&"%"'
        )
        ws[f"B{next_row}"].font = Font(bold=True, color="0000FF")
        next_row += 2

    # ---------- EXCEPTION/DEVIATION BLOCK ----------
    ws.merge_cells(f"A{next_row}:H{next_row}")
    ws[f"A{next_row}"] = "EXCEPTION / DEVIATION DOCUMENTATION (Complete if Status = Partial or Non-Compliant)"
    ws[f"A{next_row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{next_row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{next_row}"].alignment = Alignment(horizontal="center", vertical="center")

    next_row += 1
    exception_fields = [
        ("Formal exception request submitted:", "Yes,No"),
        ("Exception ID:", ""),
        ("Risk acceptance documented:", "Yes,No"),
        ("Risk ID:", ""),
        ("Compensating Controls (summary):", ""),
        ("☐ Network segmentation / firewall restrictions", ""),
        ("☐ Enhanced monitoring / alerting", ""),
        ("☐ IP allowlisting / access restrictions", ""),
        ("☐ Other (describe):", ""),
        ("Remediation actions required:", ""),
        ("Responsible person:", ""),
        ("Target completion date:", ""),
        ("Budget required:", "Yes,No,Unknown"),
    ]

    for label, options in exception_fields:
        ws[f"A{next_row}"] = label
        ws[f"A{next_row}"].font = Font(bold=True)
        ws.merge_cells(f"B{next_row}:D{next_row}")
        ws[f"B{next_row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{next_row}"].border = styles["border"]
        ws[f"B{next_row}"].alignment = styles["input_cell"]["alignment"]

        if options:
            dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=False)
            ws.add_data_validation(dv)
            dv.add(ws[f"B{next_row}"])

        next_row += 1

    # ---------- NOTES ----------
    next_row += 1
    ws.merge_cells(f"A{next_row}:H{next_row}")
    ws[f"A{next_row}"] = "ADDITIONAL NOTES / COMMENTS"
    ws[f"A{next_row}"].font = Font(bold=True, size=11)

    next_row += 1
    ws.merge_cells(f"A{next_row}:H{next_row+5}")
    ws[f"A{next_row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{next_row}"].border = styles["border"]
    ws[f"A{next_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    return (start_row, end_row, status_col_idx)


# ============================================================================
# END OF PART 2
# ============================================================================

# ============================================================================
# SECTION 5: INDIVIDUAL ASSESSMENT SHEET DEFINITIONS
# ============================================================================

def create_1_1_external_https_tls(ws, styles):
    """1.1 External HTTPS/TLS - matches markdown spec exactly."""
    checklist = [
        "TLS 1.3 preferred OR TLS 1.2 minimum",
        "Valid certificates from trusted public CA",
        "Certificate validity ≤ 397 days",
        "Self-signed certificates NOT used in production",
        "HTTP automatically redirects to HTTPS",
        "HSTS header configured",
        "Strong cipher suites configured",
        "Weak protocols disabled (TLS 1.1, 1.0, SSL)",
        "Perfect Forward Secrecy (PFS) enabled",
        "Certificate expiration alerts (≥ 30 days)",
    ]
    
    additional_details = [
        ("Number of external web services:", ""),
        ("Certificate expiration monitoring configured:", "Yes,No"),
        ("Automated certificate renewal:", "Yes,No,Planned"),
        ("Certificate inventory maintained:", "Yes,No"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="1.1 HTTPS/TLS - External Web Services",
        policy_ref="All externally accessible web services MUST use TLS encryption with valid certificates from trusted CAs (Policy Section 2.2.1)",
        question="Does your organization have external-facing web services or websites?",
        section_key="1.1_external_https",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_1_2_internal_https_tls(ws, styles):
    """1.2 Internal HTTPS/TLS - matches markdown spec."""
    checklist = [
        "TLS 1.2+ for services with Confidential/Restricted data",
        "Valid certificates (internal CA acceptable)",
        "Services with non-sensitive data: Risk assessed and documented",
        "Certificate inventory maintained",
        "Certificate expiration monitoring configured",
    ]
    
    additional_details = [
        ("Number of internal web services:", ""),
        ("Data classification(s) handled:", "Public,Internal,Confidential,Restricted"),
        ("Internal CA in use:", "Yes,No"),
        ("Certificate management process documented:", "Yes,No"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="1.2 HTTPS/TLS - Internal Web Services",
        policy_ref="Internal web services containing sensitive data (Confidential or Restricted classification) MUST use TLS (Policy Section 2.2.1)",
        question="Does your organization have internal web services (intranet, internal portals, internal APIs)?",
        section_key="1.2_internal_https",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_2_1_email_encryption(ws, styles):
    """2.1 Email Encryption - matches markdown spec."""
    checklist = [
        "S/MIME or PGP/GPG encryption available",
        "Users trained on when/how to encrypt sensitive emails",
        "PKI infrastructure in place for S/MIME",
        "Opportunistic TLS enabled for mail server connections",
        "STARTTLS enabled for SMTP",
    ]
    
    additional_details = [
        ("Encryption solution in use:", "S/MIME,PGP/GPG,TLS-only,Other"),
        ("PKI infrastructure for email:", "Implemented,Planned,Not implemented"),
        ("User training provided:", "Yes,No"),
        ("Encryption enforced via policy/DLP:", "Yes,No"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="2.1 Email Encryption",
        policy_ref="Emails containing classified information (Confidential or Restricted) MUST be encrypted when sent externally (Policy Section 2.2.2)",
        question="Does your organization send emails containing classified information to external parties?",
        section_key="2.1_email_encryption",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_2_2_digital_signatures(ws, styles):
    """2.2 Digital Signatures - matches markdown spec."""
    checklist = [
        "Digital signatures available for required use cases",
        "Email certificates issued from trusted CA or internal PKI",
        "Certificate validity ≤ 1 year",
        "Users trained on digital signature usage",
    ]
    
    additional_details = [
        ("Digital signatures used for:", "All emails,Legal docs only,Executives only,Not used"),
        ("Certificate source:", "Public CA,Internal PKI,Both"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="2.2 Digital Signatures",
        policy_ref="Digital signatures RECOMMENDED for all external emails, REQUIRED for legal/financial/official communications (Policy Section 2.2.2)",
        question="Does your organization use digital signatures for email?",
        section_key="2.2_digital_signatures",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_3_1_file_transfer(ws, styles):
    """3.1 File Transfer Protocols - matches markdown spec."""
    checklist = [
        "SFTP, FTPS, or HTTPS used for sensitive file transfers",
        "Plain FTP NOT used for sensitive data",
        "Strong authentication configured (key-based preferred)",
        "MFA required for external file transfer",
        "File transfer logging enabled",
        "SSH keys rotated annually (if SFTP used)",
    ]
    
    additional_details = [
        ("Approved protocols in use:", "SFTP,FTPS,HTTPS,SCP,Other"),
        ("Prohibited protocols detected:", "FTP,TFTP,None"),
        ("Authentication:", "Password,Key-based,MFA,Certificate"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="3.1 Secure File Transfer",
        policy_ref="File transfers containing sensitive data MUST use encrypted protocols (SFTP, FTPS, HTTPS). Unencrypted FTP is PROHIBITED (Policy Section 2.2.3)",
        question="Does your organization transfer files containing sensitive data to/from external parties or between systems?",
        section_key="3.1_file_transfer",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_4_1_vpn(ws, styles):
    """4.1 VPN - matches markdown spec."""
    checklist = [
        "Approved VPN protocol (IPsec/IKEv2, WireGuard, OpenVPN with TLS 1.2+)",
        "AES-256 or ChaCha20 encryption",
        "Perfect Forward Secrecy (PFS) enabled",
        "MFA required for all VPN connections",
        "Certificate-based authentication (preferred) OR strong pre-shared key",
        "VPN session timeout configured (≤30 minutes idle)",
        "Split-tunneling disabled (or documented exception)",
        "VPN access logs retained and reviewed",
    ]
    
    additional_details = [
        ("VPN protocol in use:", "IPsec/IKEv2,WireGuard,OpenVPN,Other"),
        ("Encryption algorithm:", ""),
        ("Number of VPN users:", ""),
        ("MFA solution:", "TOTP,Push notification,SMS,Hardware token,None"),
        ("Split-tunneling:", "Disabled,Enabled (with justification)"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="4.1 VPN (Virtual Private Network)",
        policy_ref="All remote access to organizational networks MUST use encrypted VPN with approved protocols (IPsec, WireGuard, OpenVPN). MFA REQUIRED (Policy Section 2.2.4)",
        question="Does your organization provide remote access via VPN?",
        section_key="4.1_vpn",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_4_2_ssh(ws, styles):
    """4.2 SSH - matches markdown spec."""
    checklist = [
        "SSH protocol version 2 only (SSHv1 disabled)",
        "Key-based authentication (password auth disabled preferred)",
        "Minimum key length: RSA 2048-bit or Ed25519",
        "Root login disabled",
        "SSH keys rotated annually",
        "Unused SSH keys removed",
        "Strong algorithms configured (per Policy Appendix A)",
        "Weak algorithms disabled (DSA, MD5, SHA-1)",
    ]
    
    additional_details = [
        ("SSH protocol version:", "SSHv2 only,SSHv1 (legacy - prohibited)"),
        ("Authentication method:", "Key-based only,Password allowed,Both"),
        ("SSH key types in use:", "Ed25519,RSA 3072+,RSA 2048,ECDSA,Other"),
        ("Root login via SSH:", "Disabled,Enabled"),
        ("SSH key rotation schedule:", "Annual,On personnel change,No rotation"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="4.2 SSH (Secure Shell)",
        policy_ref="SSH REQUIRED for all administrative and remote terminal access. SSH protocol version 2 REQUIRED, password authentication SHOULD be disabled (Policy Section 2.2.4)",
        question="Does your organization use SSH for remote system administration?",
        section_key="4.2_ssh",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_4_3_rdp(ws, styles):
    """4.3 RDP - matches markdown spec."""
    checklist = [
        "RDP accessed through VPN or jump host (NOT directly exposed)",
        "TLS encryption configured",
        "Network Level Authentication (NLA) enabled",
        "RDP encryption level set to 'High'",
        "MFA required for production system access",
        "RDP session recording (recommended for privileged access)",
    ]
    
    additional_details = [
        ("RDP access method:", "VPN required,Jump host/bastion,Direct (prohibited),Zero-trust gateway"),
        ("Network Level Authentication (NLA):", "Enabled,Disabled"),
        ("MFA for RDP:", "Required,Optional,Not implemented"),
        ("RDP encryption level:", "High,Medium,Low"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="4.3 RDP (Remote Desktop Protocol)",
        policy_ref="RDP connections MUST be encrypted using TLS. RDP MUST NOT be directly exposed to the Internet (Policy Section 2.2.4)",
        question="Does your organization use RDP for remote access?",
        section_key="4.3_rdp",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_5_1_api_security(ws, styles):
    """5.1 API Security - matches markdown spec."""
    checklist = [
        "All API endpoints use HTTPS (TLS 1.2+)",
        "HTTP endpoints disabled or blocked",
        "OAuth 2.0 with PKCE (for user-facing APIs)",
        "API keys have ≥256-bit entropy",
        "API keys rotate every 90 days",
        "API keys stored in secrets manager (NOT in code/config)",
        "Access tokens expire within 1 hour",
        "Refresh tokens expire within 24 hours",
        "Rate limiting implemented per API key",
        "API keys NOT passed in URL query parameters",
    ]
    
    additional_details = [
        ("Number of APIs:", ""),
        ("API authentication:", "OAuth 2.0,API keys,Mutual TLS,JWT,Basic auth,None"),
        ("API key entropy:", "≥256 bits,<256 bits,Unknown"),
        ("API key rotation:", "Automated (90 days),Manual,No rotation"),
        ("API keys stored in:", "Secrets manager,Environment variables,Code (prohibited),Config files"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="5.1 API Security",
        policy_ref="API endpoints MUST use HTTPS with TLS 1.2+. API authentication MUST use OAuth 2.0, API keys (256-bit entropy), or mutual TLS (Policy Section 2.2.5)",
        question="Does your organization expose APIs (REST, SOAP, GraphQL, etc.)?",
        section_key="5.1_api",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_6_1_database_connections(ws, styles):
    """6.1 Database Connections - matches markdown spec."""
    checklist = [
        "Database connections encrypted (TLS/SSL)",
        "Certificate validation enabled (not 'trust any certificate')",
        "Self-signed certificates only for internal databases (with proper CA)",
        "Unencrypted connections disabled or documented exception",
    ]
    
    additional_details = [
        ("Database systems in use:", "PostgreSQL,MySQL,MSSQL,Oracle,MongoDB,Other"),
        ("Connection encryption enforced:", "Yes,No,Mixed"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="6.1 Database Connections",
        policy_ref="Database connections MUST use encrypted protocols (TLS for PostgreSQL/MySQL, encrypted connections for MSSQL) (Policy Section 2.2.6)",
        question="Does your organization have applications connecting to databases?",
        section_key="6.1_database",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_6_2_wireless_networks(ws, styles):
    """6.2 Wireless Networks - matches markdown spec."""
    checklist = [
        "WPA3-Enterprise or WPA2-Enterprise for corporate networks",
        "802.1X with EAP-TLS (certificate-based) preferred",
        "WPA2-Personal only with strong passphrase (≥20 characters)",
        "WEP and WPA (original) NOT used",
        "Guest wireless isolated from corporate network",
        "WiFi passwords rotated quarterly (for PSK networks)",
    ]
    
    additional_details = [
        ("Corporate WiFi encryption:", "WPA3-Enterprise,WPA2-Enterprise,WPA2-Personal,Other"),
        ("Guest WiFi:", "Isolated network,Captive portal,Open (no encryption)"),
        ("802.1X authentication:", "Implemented,Planned,Not implemented"),
        ("WiFi password strength:", "≥20 characters,<20 characters"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="6.2 Wireless Networks",
        policy_ref="Wireless networks MUST use WPA3-Enterprise or WPA2-Enterprise minimum. WEP and WPA (original) PROHIBITED (Policy Section 2.2.6)",
        question="Does your organization have wireless networks?",
        section_key="6.2_wireless",
        checklist_items=checklist,
        additional_details=additional_details,
    )


def create_7_1_cloud_transmission(ws, styles):
    """7.1 Cloud Data Transmission - matches markdown spec."""
    checklist = [
        "TLS 1.2+ for all cloud API connections",
        "Private connectivity for Confidential/Restricted data (preferred)",
        "Cloud provider native encryption enabled",
        "Data encrypted in transit to/from cloud",
    ]
    
    additional_details = [
        ("Connection type:", "Public internet (TLS),VPN,Private Link/PrivateLink,Direct Connect/ExpressRoute"),
        ("Data classification transmitted:", "Public,Internal,Confidential,Restricted"),
        ("Cloud providers used:", "AWS,Azure,GCP,Office 365,Salesforce,Other"),
    ]
    
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="7.1 Cloud Data Transmission",
        policy_ref="Connections to cloud provider APIs MUST use TLS 1.2+. Private connectivity options (PrivateLink, Private Link, Private Service Connect) RECOMMENDED for high-volume or sensitive data (Policy Section 2.2.7)",
        question="Does your organization use cloud services (AWS, Azure, GCP, SaaS)?",
        section_key="7.1_cloud",
        checklist_items=checklist,
        additional_details=additional_details,
    )


# ============================================================================
# END OF PART 3
# ============================================================================

# ============================================================================
# SECTION 6: SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard(ws, styles):
    """Create summary dashboard with compliance statistics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "DATA TRANSMISSION ASSESSMENT - COMPLIANCE SUMMARY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Map dashboard rows to sheet names and their status column positions
    areas = [
        ("1.1 External HTTPS/TLS", "1.1 External HTTPS-TLS", 5),
        ("1.2 Internal HTTPS/TLS", "1.2 Internal HTTPS-TLS", 5),
        ("2.1 Email Encryption", "2.1 Email Encryption", 5),
        ("2.2 Digital Signatures", "2.2 Digital Signatures", 4),
        ("3.1 File Transfer", "3.1 File Transfer Protocols", 5),
        ("4.1 VPN", "4.1 VPN", 5),
        ("4.2 SSH", "4.2 SSH", 5),
        ("4.3 RDP", "4.3 RDP", 5),
        ("5.1 API Security", "5.1 API Security", 6),
        ("6.1 Database Connections", "6.1 Database Connections", 4),
        ("6.2 Wireless Networks", "6.2 Wireless Networks", 4),
        ("7.1 Cloud Transmission", "7.1 Cloud Transmission", 4),
    ]

    row += 1
    start_data_row = row
    for label, sheet, status_col in areas:
        ws.cell(row=row, column=1, value=label)
        
        status_col_letter = get_column_letter(status_col)
        status_range = f"'{sheet}'!{status_col_letter}8:{status_col_letter}20"

        ws.cell(row=row, column=2, value=f'=COUNTA({status_range})')
        ws.cell(row=row, column=3, value=f'=COUNTIF({status_range},"✅*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({status_range},"⚠️*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({status_range},"❌*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({status_range},"N/A")')
        ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{row-1})")
        cell.font = Font(bold=True)

    total_pct = ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
    total_pct.font = Font(bold=True, color="0000FF", size=12)

    # Column widths
    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 15

    # Critical gaps section
    row += 3
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "CRITICAL GAPS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    gap_headers = ["Priority", "Assessment Area", "Gap Description", "Responsible Person", "Target Date", "Status"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Priority dropdown
    dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=False)
    ws.add_data_validation(dv_priority)

    # Status dropdown for gaps
    dv_gap_status = DataValidation(type="list", formula1='"Not started,In progress,Completed,Blocked"', allow_blank=False)
    ws.add_data_validation(dv_gap_status)

    for i in range(10):
        row += 1
        ws.cell(row=row, column=1).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=1).border = styles["border"]
        dv_priority.add(ws.cell(row=row, column=1))
        
        for col in range(2, 7):
            c = ws.cell(row=row, column=col)
            c.fill = styles["input_cell"]["fill"]
            c.border = styles["border"]
            c.alignment = styles["input_cell"]["alignment"]
        
        dv_gap_status.add(ws.cell(row=row, column=6))

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create evidence register for audit traceability."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID",
        "Assessment Area",
        "Evidence Type",
        "Description",
        "Location/Path",
        "Date Collected",
        "Collected By",
        "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Validation dropdowns
    dv_type = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Network scan,Documentation,Vendor spec,Certificate inventory,Audit log,Compliance report,Other"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_type)

    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_ver)

    # Data rows
    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EV-{r-4:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 8: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create approval and sign-off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", "ISMS-IMP-A.8.24.1 - Data Transmission Assessment"),
        ("Assessment Period", ""),
        ("Overall Compliance Rate", "='Summary Dashboard'!G15"),  # Total row
        ("Assessment Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)

        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Status dropdown
    status_cell_row = row - 1
    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires remediation,Re-assessment required"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{status_cell_row}"])

    # Assessment Completed By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    completion_fields = ["Name", "Role/Title", "Department", "Email", "Date"]
    row += 1
    for field in completion_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Reviewed By (ISO/Security Manager)
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (Information Security Officer)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    review_fields = ["Name", "Date", "Signature"]
    row += 1
    for field in review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Approved By (CISO)
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    approval_fields = ["Name", "Date", "Approval Decision", "Signature"]
    row += 1
    for field in approval_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Approval Decision dropdown (3rd field from approval section start)
    decision_cell_row = row - 2
    dv_dec = DataValidation(type="list", formula1='"Approved,Approved with conditions,Rejected"', allow_blank=False)
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{decision_cell_row}"])

    # Next Review Date
    row += 2
    ws[f"A{row}"] = "Next Review Date:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]
    ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 9: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    print("=" * 78)
    print("ISMS-IMP-A.8.24.1 - Data Transmission Assessment Generator")
    print("ISO/IEC 27001:2022 Control A.8.24: Use of Cryptography")
    print("=" * 78)

    wb = create_workbook()
    styles = setup_styles()

    print("\n[1/16] Creating Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)

    print("[2/16] Creating 1.1 External HTTPS/TLS sheet...")
    create_1_1_external_https_tls(wb["1.1 External HTTPS-TLS"], styles)

    print("[3/16] Creating 1.2 Internal HTTPS/TLS sheet...")
    create_1_2_internal_https_tls(wb["1.2 Internal HTTPS-TLS"], styles)

    print("[4/16] Creating 2.1 Email Encryption sheet...")
    create_2_1_email_encryption(wb["2.1 Email Encryption"], styles)

    print("[5/16] Creating 2.2 Digital Signatures sheet...")
    create_2_2_digital_signatures(wb["2.2 Digital Signatures"], styles)

    print("[6/16] Creating 3.1 File Transfer sheet...")
    create_3_1_file_transfer(wb["3.1 File Transfer Protocols"], styles)

    print("[7/16] Creating 4.1 VPN sheet...")
    create_4_1_vpn(wb["4.1 VPN"], styles)

    print("[8/16] Creating 4.2 SSH sheet...")
    create_4_2_ssh(wb["4.2 SSH"], styles)

    print("[9/16] Creating 4.3 RDP sheet...")
    create_4_3_rdp(wb["4.3 RDP"], styles)

    print("[10/16] Creating 5.1 API Security sheet...")
    create_5_1_api_security(wb["5.1 API Security"], styles)

    print("[11/16] Creating 6.1 Database Connections sheet...")
    create_6_1_database_connections(wb["6.1 Database Connections"], styles)

    print("[12/16] Creating 6.2 Wireless Networks sheet...")
    create_6_2_wireless_networks(wb["6.2 Wireless Networks"], styles)

    print("[13/16] Creating 7.1 Cloud Transmission sheet...")
    create_7_1_cloud_transmission(wb["7.1 Cloud Transmission"], styles)

    print("[14/16] Creating Summary Dashboard sheet...")
    create_summary_dashboard(wb["Summary Dashboard"], styles)

    print("[15/16] Creating Evidence Register sheet...")
    create_evidence_register(wb["Evidence Register"], styles)

    print("[16/16] Creating Approval Sign-Off sheet...")
    create_approval_signoff(wb["Approval Sign-Off"], styles)

    filename = f"ISMS-IMP-A.8.24.1_Data_Transmission_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    print(f"\n✅ SUCCESS: {filename}")
    print("\nNext steps:")
    print("  1) Complete document information in Instructions & Legend")
    print("  2) Fill yellow cells in each assessment sheet (1.1–7.1)")
    print("  3) Check compliance checklists per section")
    print("  4) Document exceptions/deviations as needed")
    print("  5) Maintain Evidence Register entries")
    print("  6) Review Summary Dashboard for compliance gaps")
    print("  7) Complete Approval Sign-Off")
    print("\n" + "=" * 78)


if __name__ == "__main__":
    main()


# ============================================================================
# END OF SCRIPT
# ============================================================================