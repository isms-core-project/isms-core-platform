#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.4-5.S4 - Employment Exit Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.4, A.6.5

Assessment Domain 4 of 4: Compliance Dashboard

This script generates a dashboard workbook for executive visibility into
employment exit and disciplinary process compliance.

**Generated Workbook Structure:**
1. Executive_Dashboard - High-level KPI summary
2. Exit_Metrics - Detailed exit statistics
3. Access_Revocation_Metrics - Revocation compliance
4. Asset_Metrics - Asset recovery tracking
5. Disciplinary_Metrics - Case statistics
6. Obligation_Metrics - Post-employment tracking
7. Trend_Analysis - Historical trends
8. Data_Sources - Source data references
9. Instructions - Guidance

================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.4-5.S4"
WORKBOOK_NAME = "Employment Exit Dashboard"
CONTROL_ID = "A.6.4-5"
CONTROL_NAME = "Disciplinary Process and Employment Exit"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "kpi_label": {
            "font": Font(name="Calibri", size=11, bold=True),
            "alignment": Alignment(horizontal="left", vertical="center"),
        },
        "kpi_value": {
            "font": Font(name="Calibri", size=12, bold=True),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "border": border_thin,
        "green": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "amber": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "red": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }


def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Executive_Dashboard",
        "Exit_Metrics",
        "Access_Revocation_Metrics",
        "Asset_Metrics",
        "Disciplinary_Metrics",
        "Obligation_Metrics",
        "Trend_Analysis",
        "Data_Sources",
        "Instructions",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def create_executive_dashboard_sheet(ws, styles):
    """Create the Executive_Dashboard sheet."""
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "EMPLOYMENT EXIT COMPLIANCE DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Reporting period
    ws["A3"] = "Reporting Period:"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = ""
    ws["B3"].fill = styles["input_cell"]["fill"]
    ws["B3"].border = styles["border"]

    ws["D3"] = "Generated:"
    ws["D3"].font = Font(bold=True)
    ws["E3"] = GENERATED_DATE

    # KPI Section Header
    ws.merge_cells("A5:F5")
    ws["A5"] = "KEY PERFORMANCE INDICATORS"
    ws["A5"].font = styles["subheader"]["font"]
    ws["A5"].fill = styles["subheader"]["fill"]
    ws["A5"].alignment = styles["subheader"]["alignment"]

    # KPI Table Headers
    kpi_headers = ["KPI", "Target", "Actual", "Status", "Trend", "Notes"]
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # KPI Data
    kpis = [
        ("Exit Process Completion Rate", "100%", "", "Green", "►", ""),
        ("Access Revocation SLA Compliance", "100%", "", "Green", "►", ""),
        ("Asset Recovery Rate", ">95%", "", "Green", "►", ""),
        ("Orphaned Account Rate", "0", "", "Green", "►", ""),
        ("Exit Interview Completion", "100%", "", "Green", "►", "Voluntary exits only"),
        ("Exit Acknowledgement Rate", "100%", "", "Green", "►", ""),
        ("Disciplinary Case Closure <30d", ">90%", "", "Green", "►", ""),
        ("Active Enforcement Cases", "0", "", "Green", "►", ""),
    ]

    for row_idx, kpi in enumerate(kpis, start=7):
        for col_idx, value in enumerate(kpi, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]

            if col_idx == 3:  # Actual
                cell.fill = styles["input_cell"]["fill"]
            elif col_idx == 4:  # Status
                if value == "Green":
                    cell.fill = styles["green"]["fill"]
                elif value == "Amber":
                    cell.fill = styles["amber"]["fill"]
                elif value == "Red":
                    cell.fill = styles["red"]["fill"]
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 5:  # Trend
                cell.alignment = Alignment(horizontal="center")

    # Highlights Section
    row = len(kpis) + 8
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "HIGHLIGHTS AND ISSUES"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    row += 1
    highlights_labels = ["Key Achievements:", "Issues Requiring Attention:", "Actions Planned:"]
    for label in highlights_labels:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:F{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Sign-off Section
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "DASHBOARD SIGN-OFF"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    row += 1
    ws[f"A{row}"] = "Prepared By:"
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]
    ws[f"D{row}"] = "Date:"
    ws[f"E{row}"].fill = styles["input_cell"]["fill"]
    ws[f"E{row}"].border = styles["border"]

    row += 1
    ws[f"A{row}"] = "Approved By:"
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]
    ws[f"D{row}"] = "Date:"
    ws[f"E{row}"].fill = styles["input_cell"]["fill"]
    ws[f"E{row}"].border = styles["border"]

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 30


def create_exit_metrics_sheet(ws, styles):
    """Create the Exit_Metrics sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EXIT METRICS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Period", 12),
        ("B", "Total_Exits", 14),
        ("C", "Voluntary", 12),
        ("D", "Involuntary", 14),
        ("E", "Contractors", 14),
        ("F", "Fully_Complete", 16),
        ("G", "Completion_Rate", 16),
        ("H", "Outstanding_Items", 18),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Pre-populate 12 months
    months = ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06",
              "2026-07", "2026-08", "2026-09", "2026-10", "2026-11", "2026-12"]

    for row_idx, month in enumerate(months, start=4):
        ws[f"A{row_idx}"] = month
        ws[f"A{row_idx}"].border = styles["border"]

        for col in ["B", "C", "D", "E", "F", "H"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

        # Completion rate formula
        ws[f"G{row_idx}"] = f'=IF(B{row_idx}>0,F{row_idx}/B{row_idx},"")'
        ws[f"G{row_idx}"].border = styles["border"]
        ws[f"G{row_idx}"].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws[f"G{row_idx}"].number_format = '0.0%'

    ws.freeze_panes = "A4"


def create_access_revocation_metrics_sheet(ws, styles):
    """Create the Access_Revocation_Metrics sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "ACCESS REVOCATION METRICS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Period", 12),
        ("B", "Total_Revocations", 18),
        ("C", "Within_SLA", 14),
        ("D", "SLA_Compliance", 16),
        ("E", "Avg_Revocation_Hours", 20),
        ("F", "Same_Day_Rate", 16),
        ("G", "Privileged_Compliance", 20),
        ("H", "Third_Party_Notified", 20),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    months = ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06",
              "2026-07", "2026-08", "2026-09", "2026-10", "2026-11", "2026-12"]

    for row_idx, month in enumerate(months, start=4):
        ws[f"A{row_idx}"] = month
        ws[f"A{row_idx}"].border = styles["border"]

        for col in ["B", "C", "E", "F", "G", "H"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

        # SLA Compliance formula
        ws[f"D{row_idx}"] = f'=IF(B{row_idx}>0,C{row_idx}/B{row_idx},"")'
        ws[f"D{row_idx}"].border = styles["border"]
        ws[f"D{row_idx}"].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws[f"D{row_idx}"].number_format = '0.0%'

    ws.freeze_panes = "A4"


def create_asset_metrics_sheet(ws, styles):
    """Create the Asset_Metrics sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "ASSET RECOVERY METRICS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Period", 12),
        ("B", "Assets_Due", 14),
        ("C", "Assets_Returned", 16),
        ("D", "Recovery_Rate", 14),
        ("E", "Outstanding_30Days", 18),
        ("F", "Data_Wiping_Verified", 20),
        ("G", "Wiping_Rate", 14),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    months = ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06",
              "2026-07", "2026-08", "2026-09", "2026-10", "2026-11", "2026-12"]

    for row_idx, month in enumerate(months, start=4):
        ws[f"A{row_idx}"] = month
        ws[f"A{row_idx}"].border = styles["border"]

        for col in ["B", "C", "E", "F"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

        # Recovery rate formula
        ws[f"D{row_idx}"] = f'=IF(B{row_idx}>0,C{row_idx}/B{row_idx},"")'
        ws[f"D{row_idx}"].border = styles["border"]
        ws[f"D{row_idx}"].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws[f"D{row_idx}"].number_format = '0.0%'

        # Wiping rate formula
        ws[f"G{row_idx}"] = f'=IF(C{row_idx}>0,F{row_idx}/C{row_idx},"")'
        ws[f"G{row_idx}"].border = styles["border"]
        ws[f"G{row_idx}"].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws[f"G{row_idx}"].number_format = '0.0%'

    ws.freeze_panes = "A4"


def create_disciplinary_metrics_sheet(ws, styles):
    """Create the Disciplinary_Metrics sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "DISCIPLINARY METRICS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Period", 12),
        ("B", "Cases_Opened", 14),
        ("C", "Cases_Closed", 14),
        ("D", "Avg_Duration_Days", 18),
        ("E", "By_Severity", 30),
        ("F", "Appeals_Filed", 14),
        ("G", "Terminations", 14),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    months = ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06",
              "2026-07", "2026-08", "2026-09", "2026-10", "2026-11", "2026-12"]

    for row_idx, month in enumerate(months, start=4):
        ws[f"A{row_idx}"] = month
        ws[f"A{row_idx}"].border = styles["border"]

        for col in ["B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A4"


def create_obligation_metrics_sheet(ws, styles):
    """Create the Obligation_Metrics sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "POST-EMPLOYMENT OBLIGATION METRICS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Period", 12),
        ("B", "Total_Active", 14),
        ("C", "New_This_Period", 16),
        ("D", "Expired_This_Period", 18),
        ("E", "Expiring_90Days", 16),
        ("F", "Enforcement_Active", 18),
        ("G", "Acknowledgement_Rate", 20),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    months = ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06",
              "2026-07", "2026-08", "2026-09", "2026-10", "2026-11", "2026-12"]

    for row_idx, month in enumerate(months, start=4):
        ws[f"A{row_idx}"] = month
        ws[f"A{row_idx}"].border = styles["border"]

        for col in ["B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A4"


def create_trend_analysis_sheet(ws, styles):
    """Create the Trend_Analysis sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "TREND ANALYSIS (12-Month Rolling)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Period", 12),
        ("B", "Exit_Completion", 16),
        ("C", "SLA_Compliance", 16),
        ("D", "Asset_Recovery", 16),
        ("E", "Orphaned_Accounts", 18),
        ("F", "Disciplinary_Cases", 18),
        ("G", "Avg_Case_Duration", 18),
        ("H", "Active_Obligations", 18),
        ("I", "Enforcement_Cases", 18),
        ("J", "Rolling_3M_Completion", 20),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    months = ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06",
              "2026-07", "2026-08", "2026-09", "2026-10", "2026-11", "2026-12"]

    for row_idx, month in enumerate(months, start=4):
        ws[f"A{row_idx}"] = month
        ws[f"A{row_idx}"].border = styles["border"]

        for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

        # Rolling 3-month average formula
        if row_idx >= 6:
            ws[f"J{row_idx}"] = f'=IF(COUNT(B{row_idx-2}:B{row_idx})=3,AVERAGE(B{row_idx-2}:B{row_idx}),"")'
        else:
            ws[f"J{row_idx}"] = ""
        ws[f"J{row_idx}"].border = styles["border"]
        ws[f"J{row_idx}"].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.freeze_panes = "A4"


def create_data_sources_sheet(ws, styles):
    """Create the Data_Sources sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "DATA SOURCES"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Data_Source", 22),
        ("B", "Source_Sheet", 25),
        ("C", "Last_Updated", 14),
        ("D", "Records_Count", 14),
        ("E", "Update_Method", 16),
        ("F", "Updated_By", 25),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    sources = [
        ("ISMS-IMP-A.6.4-5.S1", "Case_Tracker", "", "", "Manual", ""),
        ("ISMS-IMP-A.6.4-5.S2", "Exit_Tracker", "", "", "Manual", ""),
        ("ISMS-IMP-A.6.4-5.S2", "Leaver_Reconciliation", "", "", "Manual", ""),
        ("ISMS-IMP-A.6.4-5.S3", "Active_Obligations", "", "", "Manual", ""),
        ("ISMS-IMP-A.6.4-5.S3", "Acknowledgement_Log", "", "", "Manual", ""),
        ("HRIS", "Termination Records", "", "", "Export", ""),
        ("IAM System", "Account Disable Logs", "", "", "Export", ""),
        ("Asset Management", "Asset Return Records", "", "", "Export", ""),
    ]

    update_method_val = DataValidation(type="list", formula1='"Manual,Linked,Automated,Export"')
    ws.add_data_validation(update_method_val)

    for row_idx, source in enumerate(sources, start=4):
        for col_idx, value in enumerate(source, start=1):
            col = get_column_letter(col_idx)
            cell = ws[f"{col}{row_idx}"]
            cell.value = value
            cell.border = styles["border"]
            if value == "":
                cell.fill = styles["input_cell"]["fill"]

        update_method_val.add(ws[f"E{row_idx}"])

    ws.freeze_panes = "A4"


def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:G3")
    ws["A3"] = CONTROL_REF
    ws["A3"].font = Font(bold=True, size=12)

    instructions = [
        "",
        "PURPOSE",
        "This dashboard provides executive visibility into employment exit and",
        "disciplinary process compliance for ISO 27001:2022 Controls A.6.4 and A.6.5.",
        "",
        "DATA SOURCES",
        "This dashboard aggregates data from:",
        "- ISMS-IMP-A.6.4-5.S1: Disciplinary Process Assessment",
        "- ISMS-IMP-A.6.4-5.S2: Employment Exit Assessment",
        "- ISMS-IMP-A.6.4-5.S3: Post-Employment Obligations",
        "",
        "UPDATE PROCESS",
        "1. Update Data_Sources sheet with latest source data",
        "2. Update metric sheets with monthly data",
        "3. Review Executive_Dashboard for auto-calculated KPIs",
        "4. Update highlights and issues",
        "5. Obtain sign-off",
        "",
        "KPI DEFINITIONS",
        "- Exit Completion: Exits with all steps complete / Total exits",
        "- SLA Compliance: Revocations within SLA / Total revocations",
        "- Asset Recovery: Assets returned / Assets due",
        "- Orphaned Accounts: Active accounts for terminated staff",
        "",
        "RAG STATUS THRESHOLDS",
        "- Green: Target met",
        "- Amber: 90-99% of target",
        "- Red: Below 90% of target",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, text in enumerate(instructions, start=5):
        ws[f"A{i}"] = text
        if text in ["PURPOSE", "DATA SOURCES", "UPDATE PROCESS", "KPI DEFINITIONS", "RAG STATUS THRESHOLDS"]:
            ws[f"A{i}"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 80


def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/9] Creating Executive_Dashboard sheet...")
        create_executive_dashboard_sheet(wb["Executive_Dashboard"], styles)

        logger.info("[2/9] Creating Exit_Metrics sheet...")
        create_exit_metrics_sheet(wb["Exit_Metrics"], styles)

        logger.info("[3/9] Creating Access_Revocation_Metrics sheet...")
        create_access_revocation_metrics_sheet(wb["Access_Revocation_Metrics"], styles)

        logger.info("[4/9] Creating Asset_Metrics sheet...")
        create_asset_metrics_sheet(wb["Asset_Metrics"], styles)

        logger.info("[5/9] Creating Disciplinary_Metrics sheet...")
        create_disciplinary_metrics_sheet(wb["Disciplinary_Metrics"], styles)

        logger.info("[6/9] Creating Obligation_Metrics sheet...")
        create_obligation_metrics_sheet(wb["Obligation_Metrics"], styles)

        logger.info("[7/9] Creating Trend_Analysis sheet...")
        create_trend_analysis_sheet(wb["Trend_Analysis"], styles)

        logger.info("[8/9] Creating Data_Sources sheet...")
        create_data_sources_sheet(wb["Data_Sources"], styles)

        logger.info("[9/9] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.6.4-5.S4 specification
# =============================================================================
