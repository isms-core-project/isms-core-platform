#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.4-5.S3 - Post-Employment Obligations Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.5: Responsibilities After Termination

Assessment Domain 3 of 4: Post-Employment Obligations Tracking

This script generates a workbook for tracking and managing post-employment
obligations including NDA compliance, confidentiality, and enforcement.

**Generated Workbook Structure:**
1. Instructions - Assessment guidance
2. Obligation_Types - Obligation definitions
3. Former_Personnel - Former staff registry
4. Active_Obligations - Current active obligations
5. Expiration_Tracker - Expiration monitoring
6. Acknowledgement_Log - Exit acknowledgements
7. Enforcement_Register - Enforcement records
8. Evidence_Register - Audit evidence tracking
9. Approval_SignOff - Stakeholder approval

================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.4-5.S3"
WORKBOOK_NAME = "Post Employment Obligations"
CONTROL_ID = "A.6.4-5"
CONTROL_NAME = "Disciplinary Process and Employment Exit"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "active": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "expired": {"fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")},
        "enforcement": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "expiring": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
    }


def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Obligation_Types",
        "Former_Personnel",
        "Active_Obligations",
        "Expiration_Tracker",
        "Acknowledgement_Log",
        "Enforcement_Register",
        "Evidence_Register",
        "Approval_SignOff",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:G3")
    ws["A3"] = CONTROL_REF
    ws["A3"].font = Font(bold=True, size=12)

    instructions = [
        "",
        "PURPOSE",
        "This workbook tracks post-employment obligations for former personnel,",
        "ensuring continuing confidentiality and security responsibilities are managed.",
        "",
        "SCOPE",
        "- Post-employment confidentiality obligations (NDAs)",
        "- Intellectual property restrictions",
        "- Non-compete and non-solicitation tracking",
        "- Obligation expiration monitoring",
        "- Enforcement action records",
        "",
        "OBLIGATION TYPES",
        "- Confidentiality: Protection of non-public information (2-5 years)",
        "- Trade Secrets: Indefinite protection",
        "- IP Assignment: Perpetual organisation ownership",
        "- Non-Compete: Time-limited competitive restrictions",
        "- Non-Solicitation: Customer/employee solicitation restrictions",
        "- Data Return: Confirmation of data return/destruction",
        "",
        "COMPLETION STEPS",
        "1. Obligation_Types: Define obligation categories",
        "2. Former_Personnel: Register former staff with obligations",
        "3. Active_Obligations: Track individual obligations",
        "4. Expiration_Tracker: Monitor approaching expirations",
        "5. Acknowledgement_Log: Track exit acknowledgements",
        "6. Enforcement_Register: Document any enforcement actions",
        "7. Evidence_Register: Link supporting evidence",
        "8. Approval_SignOff: Obtain required approvals",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, text in enumerate(instructions, start=5):
        ws[f"A{i}"] = text
        if text in ["PURPOSE", "SCOPE", "OBLIGATION TYPES", "COMPLETION STEPS"]:
            ws[f"A{i}"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 80


def create_obligation_types_sheet(ws, styles):
    """Create the Obligation_Types sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "OBLIGATION TYPES"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Obligation_ID", 14),
        ("B", "Obligation_Type", 25),
        ("C", "Description", 50),
        ("D", "Standard_Duration", 20),
        ("E", "Source_Document", 30),
        ("F", "Applicable_To", 25),
        ("G", "Enforceability_Notes", 45),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Pre-populate obligation types
    obligations = [
        ("OBL-001", "Confidentiality", "Protection of all non-public information accessed during employment", "3 years", "Employment Contract, NDA", "All Employees", "Fully enforceable under Swiss law"),
        ("OBL-002", "Trade Secrets", "Protection of trade secrets and proprietary information", "Indefinite", "Employment Contract, NDA", "All Employees", "Protected under Swiss Unfair Competition Act"),
        ("OBL-003", "IP Assignment", "Work product and inventions remain organisation property", "Perpetual", "IP Assignment Agreement", "All Employees", "Strong legal protection available"),
        ("OBL-004", "Non-Compete", "Restriction on working for direct competitors", "12 months", "Employment Contract", "Executives, Key Technical", "Limited enforceability in CH; requires compensation"),
        ("OBL-005", "Non-Solicitation (Employee)", "Restriction on soliciting current employees", "24 months", "Employment Contract, NDA", "Managers, Executives", "Generally enforceable"),
        ("OBL-006", "Non-Solicitation (Customer)", "Restriction on soliciting current customers", "24 months", "Employment Contract, NDA", "Sales, Account Managers", "Generally enforceable"),
        ("OBL-007", "Data Return", "Return or certified destruction of all company data", "Immediate at exit", "NDA, Acceptable Use Policy", "All Employees", "Enforceable; obtain written certification"),
    ]

    for row_idx, obligation in enumerate(obligations, start=4):
        for col_idx, value in enumerate(obligation):
            col = get_column_letter(col_idx + 1)
            cell = ws[f"{col}{row_idx}"]
            cell.value = value
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A4"


def create_former_personnel_sheet(ws, styles):
    """Create the Former_Personnel sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "FORMER PERSONNEL REGISTRY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Person_ID", 16),
        ("B", "Name", 30),
        ("C", "Exit_Date", 14),
        ("D", "Exit_Type", 22),
        ("E", "Position_Held", 30),
        ("F", "Access_Level", 16),
        ("G", "NDA_Reference", 20),
        ("H", "Obligations_End_Date", 18),
        ("I", "Status", 18),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Data validations
    exit_type_val = DataValidation(
        type="list",
        formula1='"Voluntary Resignation,Involuntary Termination,Contract End,Retirement,Role Change"'
    )
    ws.add_data_validation(exit_type_val)

    access_val = DataValidation(
        type="list",
        formula1='"Standard,Elevated,Privileged,Executive"'
    )
    ws.add_data_validation(access_val)

    status_val = DataValidation(
        type="list",
        formula1='"Active Obligations,All Expired,Under Enforcement"'
    )
    ws.add_data_validation(status_val)

    for row in range(4, 504):
        ws[f"A{row}"] = f"FP-2025-{row-3:03d}"
        ws[f"A{row}"].font = Font(color="808080")

        for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

        exit_type_val.add(ws[f"D{row}"])
        access_val.add(ws[f"F{row}"])
        status_val.add(ws[f"I{row}"])

    ws.freeze_panes = "A4"


def create_active_obligations_sheet(ws, styles):
    """Create the Active_Obligations sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "ACTIVE OBLIGATIONS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Obligation_Ref", 18),
        ("B", "Person_ID", 16),
        ("C", "Obligation_Type", 22),
        ("D", "Start_Date", 14),
        ("E", "End_Date", 14),
        ("F", "Specific_Terms", 40),
        ("G", "Monitoring_Required", 18),
        ("H", "Status", 18),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Data validations
    obligation_type_val = DataValidation(
        type="list",
        formula1='"Confidentiality,Trade Secrets,IP Assignment,Non-Compete,Non-Solicitation (Employee),Non-Solicitation (Customer),Data Return"'
    )
    ws.add_data_validation(obligation_type_val)

    monitoring_val = DataValidation(type="list", formula1='"Yes,No,Periodic"')
    ws.add_data_validation(monitoring_val)

    status_val = DataValidation(
        type="list",
        formula1='"Active,Expired,Waived,Under Enforcement"'
    )
    ws.add_data_validation(status_val)

    for row in range(4, 1004):
        ws[f"A{row}"] = f"ACT-2025-{row-3:04d}"
        ws[f"A{row}"].font = Font(color="808080")

        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

        obligation_type_val.add(ws[f"C{row}"])
        monitoring_val.add(ws[f"G{row}"])
        status_val.add(ws[f"H{row}"])

    ws.freeze_panes = "A4"


def create_expiration_tracker_sheet(ws, styles):
    """Create the Expiration_Tracker sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EXPIRATION TRACKER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Obligation_Ref", 18),
        ("B", "Person_Name", 30),
        ("C", "Obligation_Type", 22),
        ("D", "Expiration_Date", 14),
        ("E", "Days_Until_Expiry", 18),
        ("F", "Status", 20),
        ("G", "Action_Required", 40),
        ("H", "Action_Taken", 40),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    status_val = DataValidation(
        type="list",
        formula1='"Expiring Soon,Expired,Acknowledged,Extended"'
    )
    ws.add_data_validation(status_val)

    for row in range(4, 104):
        for col in ["A", "B", "C", "D", "G", "H"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

        # Days until expiry formula
        ws[f"E{row}"] = f'=IF(D{row}="Indefinite","N/A",IF(D{row}<>"",D{row}-TODAY(),""))'
        ws[f"E{row}"].border = styles["border"]
        ws[f"E{row}"].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        ws[f"F{row}"].border = styles["border"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        status_val.add(ws[f"F{row}"])

    ws.freeze_panes = "A4"


def create_acknowledgement_log_sheet(ws, styles):
    """Create the Acknowledgement_Log sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EXIT ACKNOWLEDGEMENT LOG"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Ack_ID", 16),
        ("B", "Person_ID", 16),
        ("C", "Exit_Date", 14),
        ("D", "Acknowledgement_Type", 25),
        ("E", "Acknowledgement_Date", 18),
        ("F", "Obligations_Summarised", 20),
        ("G", "Signed_Document", 16),
        ("H", "Document_Location", 45),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    ack_type_val = DataValidation(
        type="list",
        formula1='"Exit Interview,Written Acknowledgement,Email Confirmation,Refused"'
    )
    ws.add_data_validation(ack_type_val)

    yes_no_val = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yes_no_val)

    for row in range(4, 504):
        ws[f"A{row}"] = f"ACK-2025-{row-3:03d}"
        ws[f"A{row}"].font = Font(color="808080")

        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

        ack_type_val.add(ws[f"D{row}"])
        yes_no_val.add(ws[f"F{row}"])
        yes_no_val.add(ws[f"G{row}"])

    ws.freeze_panes = "A4"


def create_enforcement_register_sheet(ws, styles):
    """Create the Enforcement_Register sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "ENFORCEMENT REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Enforcement_ID", 18),
        ("B", "Person_ID", 16),
        ("C", "Obligation_Violated", 22),
        ("D", "Violation_Date", 14),
        ("E", "Violation_Description", 45),
        ("F", "Detection_Method", 30),
        ("G", "Enforcement_Action", 40),
        ("H", "Status", 22),
        ("I", "Outcome", 40),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    obligation_val = DataValidation(
        type="list",
        formula1='"Confidentiality,Trade Secrets,IP Assignment,Non-Compete,Non-Solicitation (Employee),Non-Solicitation (Customer),Data Return"'
    )
    ws.add_data_validation(obligation_val)

    status_val = DataValidation(
        type="list",
        formula1='"Under Investigation,Cease and Desist Sent,Legal Action Initiated,Resolved,No Action Required"'
    )
    ws.add_data_validation(status_val)

    for row in range(4, 54):
        ws[f"A{row}"] = f"ENF-2026-{row-3:03d}"
        ws[f"A{row}"].font = Font(color="808080")

        for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

        obligation_val.add(ws[f"C{row}"])
        status_val.add(ws[f"H{row}"])

    ws.freeze_panes = "A4"


def create_evidence_register_sheet(ws, styles):
    """Create the Evidence_Register sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Evidence_ID", 22),
        ("B", "Evidence_Description", 50),
        ("C", "Evidence_Type", 22),
        ("D", "Storage_Location", 50),
        ("E", "Collection_Date", 14),
        ("F", "Collected_By", 25),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    type_val = DataValidation(
        type="list",
        formula1='"Obligation Definition,NDA Copy,Acknowledgement Form,Enforcement Record,Expiration Report,Other"'
    )
    ws.add_data_validation(type_val)

    for row in range(4, 54):
        ws[f"A{row}"] = f"EVD-A.6.4-5.3-{row-3:03d}"
        ws[f"A{row}"].font = Font(color="808080")

        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

        type_val.add(ws[f"C{row}"])

    ws.freeze_panes = "A4"


def create_approval_signoff_sheet(ws, styles):
    """Create the Approval_SignOff sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Obligation Types Defined", "7 (pre-populated)"),
        ("Former Personnel Tracked", "=COUNTA(Former_Personnel!B4:B503)"),
        ("Active Obligations", "=COUNTIF(Active_Obligations!H4:H1003,\"Active\")"),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "" or not str(value).startswith("="):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    row += 1
    ws[f"A{row}"] = "Role"
    ws[f"B{row}"] = "Name"
    ws[f"C{row}"] = "Signature"
    ws[f"D{row}"] = "Date"
    ws[f"E{row}"] = "Comments"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].border = styles["border"]

    reviewers = ["Legal Counsel", "HR Director", "CISO"]
    row += 1
    for role in reviewers:
        ws[f"A{row}"] = role
        ws[f"A{row}"].border = styles["border"]
        for col in ["B", "C", "D", "E"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 35


def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/9] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/9] Creating Obligation_Types sheet...")
        create_obligation_types_sheet(wb["Obligation_Types"], styles)

        logger.info("[3/9] Creating Former_Personnel sheet...")
        create_former_personnel_sheet(wb["Former_Personnel"], styles)

        logger.info("[4/9] Creating Active_Obligations sheet...")
        create_active_obligations_sheet(wb["Active_Obligations"], styles)

        logger.info("[5/9] Creating Expiration_Tracker sheet...")
        create_expiration_tracker_sheet(wb["Expiration_Tracker"], styles)

        logger.info("[6/9] Creating Acknowledgement_Log sheet...")
        create_acknowledgement_log_sheet(wb["Acknowledgement_Log"], styles)

        logger.info("[7/9] Creating Enforcement_Register sheet...")
        create_enforcement_register_sheet(wb["Enforcement_Register"], styles)

        logger.info("[8/9] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[9/9] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.6.4-5.S3 specification
# =============================================================================
