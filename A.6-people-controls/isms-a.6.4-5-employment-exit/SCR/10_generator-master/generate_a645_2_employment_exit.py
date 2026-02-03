#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.4-5.2 - Employment Exit Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.5: Responsibilities After Termination

Assessment Domain 2 of 4: Employment Exit Process

This script generates a workbook for managing employment exit processes
including access revocation, asset recovery, and exit procedures.

**Generated Workbook Structure:**
1. Instructions - Assessment guidance
2. Exit_Procedures - Exit process by type
3. Access_Revocation - Revocation requirements
4. Asset_Recovery - Asset return tracking
5. Exit_Tracker - Individual exit tracking
6. Leaver_Reconciliation - Orphaned account checks
7. Evidence_Register - Audit evidence tracking
8. Approval_SignOff - Stakeholder approval

================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.4-5.2"
WORKBOOK_NAME = "Employment Exit Assessment"
CONTROL_ID = "A.6.4-5"
CONTROL_NAME = "Disciplinary Process and Employment Exit"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
    }


def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Exit_Procedures",
        "Access_Revocation",
        "Asset_Recovery",
        "Exit_Tracker",
        "Leaver_Reconciliation",
        "Evidence_Register",
        "Approval_SignOff",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:G3")
    ws["A3"] = CONTROL_REF
    ws["A3"].font = Font(bold=True, size=12)

    instructions = [
        "",
        "PURPOSE",
        "This workbook manages employment exit processes to ensure secure offboarding",
        "of personnel with complete access revocation and asset recovery.",
        "",
        "SCOPE",
        "- Exit procedures by termination type",
        "- Access revocation timelines and verification",
        "- Asset recovery tracking",
        "- Monthly leaver reconciliation",
        "",
        "EXIT TYPES",
        "- Voluntary Resignation: Standard notice period exit",
        "- Involuntary Termination: Same-day access revocation",
        "- Immediate Dismissal: Within 1 hour, escorted exit",
        "- Retirement: Extended knowledge transfer period",
        "- Contract End: Per contract terms",
        "- Role Change: Previous role access revocation",
        "",
        "COMPLETION STEPS",
        "1. Exit_Procedures: Document processes by exit type",
        "2. Access_Revocation: Define revocation requirements",
        "3. Asset_Recovery: Specify asset return procedures",
        "4. Exit_Tracker: Track individual exits (ongoing)",
        "5. Leaver_Reconciliation: Monthly orphan account check",
        "6. Evidence_Register: Link supporting evidence",
        "7. Approval_SignOff: Obtain required approvals",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, text in enumerate(instructions, start=5):
        ws[f"A{i}"] = text
        if text in ["PURPOSE", "SCOPE", "EXIT TYPES", "COMPLETION STEPS"]:
            ws[f"A{i}"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 80


def create_exit_procedures_sheet(ws, styles):
    """Create the Exit_Procedures sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EXIT PROCEDURES BY TYPE"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Exit_Type", 25),
        ("B", "Notification_Trigger", 35),
        ("C", "HR_Actions", 45),
        ("D", "IT_Actions", 45),
        ("E", "Manager_Actions", 40),
        ("F", "Security_Actions", 40),
        ("G", "Timeline", 25),
        ("H", "Documentation_Required", 40),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Pre-populate exit procedures
    procedures = [
        ("Voluntary Resignation", "HR receives resignation letter", "Update HRIS, schedule exit interview, initiate checklist", "Disable accounts end of last day, backup data, revoke VPN", "Coordinate knowledge transfer, approve asset return", "Review access logs, verify no data exfiltration", "Per notice period (1-3 months)", "Exit checklist, asset form, interview notes"),
        ("Involuntary Termination", "Termination decision approved", "Prepare termination documents, coordinate timing with IT", "Disable accounts same day before notification, backup data", "Prepare handover, coordinate meeting room", "Immediate access log review, monitor for anomalies", "Same business day", "Termination letter, exit checklist, asset form"),
        ("Immediate Dismissal", "Gross misconduct confirmed", "Prepare immediate termination, coordinate Security escort", "Disable ALL access within 1 hour, preserve evidence", "Not involved in exit meeting", "Escort off premises, forensic preservation if needed", "Within 1 hour", "Termination letter, incident report, asset form"),
        ("Retirement", "Retirement notice received", "Update HRIS, plan recognition, extended exit interview", "Disable accounts end of last day, extended backup", "Extended knowledge transfer plan", "Standard access review", "Per notice (typically 3+ months)", "Exit checklist, asset form, knowledge transfer log"),
        ("Contract End", "Contract end date approaching", "Confirm end date, schedule exit process", "Disable accounts on contract end date", "Verify deliverables, approve asset return", "Review access for contract period", "Contract end date", "Contract closeout, asset form"),
        ("Role Change", "Internal transfer approved", "Update HRIS role, coordinate access change", "Revoke previous role access within 2 days, grant new access", "Handover responsibilities, transfer knowledge", "Verify appropriate access change", "2 business days for old access removal", "Access change form, handover notes"),
    ]

    for row_idx, procedure in enumerate(procedures, start=4):
        for col_idx, value in enumerate(procedure):
            col = get_column_letter(col_idx + 1)
            cell = ws[f"{col}{row_idx}"]
            cell.value = value
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A4"


def create_access_revocation_sheet(ws, styles):
    """Create the Access_Revocation sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "ACCESS REVOCATION REQUIREMENTS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Access_Type", 22),
        ("B", "System_Examples", 35),
        ("C", "Revocation_Method", 40),
        ("D", "Timeline_Standard", 20),
        ("E", "Timeline_Immediate", 20),
        ("F", "Verification_Method", 35),
        ("G", "Responsible_Party", 20),
        ("H", "Dependencies", 35),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Pre-populate access types
    access_types = [
        ("Physical Access", "Building badges, keys, biometrics", "Disable badge in access system, collect keys, remove biometric template", "End of last day", "Within 1 hour", "Badge log review, test badge at door", "Facilities Team", "None"),
        ("AD/Directory", "Active Directory, Azure AD, LDAP", "Disable account, move to Disabled OU, remove from groups", "End of last day", "Within 1 hour", "Login attempt test, access report", "IAM Team", "SSO-integrated apps disabled automatically"),
        ("Email", "Exchange, O365, Google Workspace", "Disable account, set auto-reply, configure forwarding if approved", "End of last day", "Within 1 hour", "Login attempt, mail flow test", "IAM Team", "Requires AD disable first"),
        ("VPN", "Corporate VPN, remote access", "Revoke VPN certificate, disable VPN account", "End of last day", "Within 1 hour", "Connection attempt test", "Network Team", "May require AD disable"),
        ("Applications", "Business apps, SaaS (non-SSO)", "Disable/remove user from each application", "End of last day + 24h", "Same day", "Login attempt per app", "Application Owners", "Coordinate with app owners"),
        ("Cloud Services", "AWS, Azure, GCP console access", "Remove from IAM, revoke API keys, remove from tenants", "End of last day", "Within 1 hour", "Console login test, API call test", "Cloud Team", "Audit CloudTrail/activity logs"),
        ("Third-Party Systems", "Partner portals, customer systems", "Notify third parties, request account removal", "Within 48 hours", "Within 24 hours", "Confirmation from third party", "Business Owner", "Requires external coordination"),
        ("Privileged Access", "Admin accounts, PAM-managed access", "Immediate revocation, rotate shared passwords", "End of last day", "Immediate (within 15 min)", "Privileged access report, login test", "Security Team", "Check PAM session recordings"),
        ("Shared Accounts", "Service accounts, shared mailboxes", "Remove delegated access, rotate passwords if needed", "End of last day", "Same day", "Access attempt, delegation list review", "IAM Team", "Document all shared access"),
        ("API Keys", "Developer keys, service credentials", "Revoke keys, rotate if shared", "End of last day", "Immediate", "API call test with revoked key", "DevOps Team", "Check for hardcoded keys"),
    ]

    for row_idx, access in enumerate(access_types, start=4):
        for col_idx, value in enumerate(access):
            col = get_column_letter(col_idx + 1)
            cell = ws[f"{col}{row_idx}"]
            cell.value = value
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A4"


def create_asset_recovery_sheet(ws, styles):
    """Create the Asset_Recovery sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "ASSET RECOVERY REQUIREMENTS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Asset_Category", 22),
        ("B", "Asset_Examples", 35),
        ("C", "Return_Method", 35),
        ("D", "Verification_Required", 35),
        ("E", "Data_Handling", 40),
        ("F", "Timeline", 20),
        ("G", "Non_Return_Action", 40),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Pre-populate asset categories
    assets = [
        ("Computing Devices", "Laptop, desktop, monitors, docking stations", "In-person handover to IT", "Serial number match, condition assessment", "Secure erase or reimage before reissue", "By last working day", "Manager follow-up, payroll deduction, report as lost"),
        ("Mobile Devices", "Company phone, tablet", "In-person handover to IT", "IMEI/serial match, condition check", "Factory reset, verify MDM unenrollment", "By last working day", "Remote wipe via MDM, report as lost"),
        ("Storage Media", "USB drives, external drives, backup media", "In-person handover to IT", "Inventory match", "Secure erase or physical destruction", "By last working day", "Report as security incident if sensitive data"),
        ("Access Devices", "Badges, keys, tokens, smart cards", "In-person handover to Facilities/Security", "Inventory match, badge ID verification", "Disable in access system, revoke certificates", "By last working day", "Disable immediately, rekey if high security"),
        ("Documentation", "Printed documents, manuals, customer materials", "In-person handover to manager", "Spot check completeness", "Secure destruction of confidential materials", "By last working day", "Manager verification, log incomplete return"),
        ("Software Licenses", "Named user licenses, dongles", "IT collection or verification of removal", "License deactivation confirmed", "Deactivate license, update inventory", "By last working day", "Deactivate remotely, reallocate license"),
        ("BYOD (if MDM)", "Personal devices with company data", "MDM unenrollment by IT", "Confirm MDM profile removed", "Selective wipe of company data via MDM", "By last working day", "Remote selective wipe"),
    ]

    for row_idx, asset in enumerate(assets, start=4):
        for col_idx, value in enumerate(asset):
            col = get_column_letter(col_idx + 1)
            cell = ws[f"{col}{row_idx}"]
            cell.value = value
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A4"


def create_exit_tracker_sheet(ws, styles):
    """Create the Exit_Tracker sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "EXIT TRACKER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Exit_ID", 18),
        ("B", "Employee_ID", 15),
        ("C", "Exit_Type", 22),
        ("D", "Last_Working_Day", 16),
        ("E", "Access_Revoked_Date", 18),
        ("F", "Assets_Returned_Date", 18),
        ("G", "Exit_Interview_Date", 18),
        ("H", "Checklist_Complete", 16),
        ("I", "Status", 18),
        ("J", "Notes", 40),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Data validations
    exit_type_val = DataValidation(
        type="list",
        formula1='"Voluntary Resignation,Involuntary Termination,Immediate Dismissal,Retirement,Contract End,Role Change"'
    )
    ws.add_data_validation(exit_type_val)

    checklist_val = DataValidation(type="list", formula1='"Yes,Partial,No"')
    ws.add_data_validation(checklist_val)

    status_val = DataValidation(
        type="list",
        formula1='"Initiated,In Progress,Pending Asset Return,Pending Verification,Complete,Issues Outstanding"'
    )
    ws.add_data_validation(status_val)

    for row in range(4, 204):
        ws[f"A{row}"] = f"EXIT-2026-{row-3:03d}"
        ws[f"A{row}"].font = Font(color="808080")

        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

        exit_type_val.add(ws[f"C{row}"])
        checklist_val.add(ws[f"H{row}"])
        status_val.add(ws[f"I{row}"])

    ws.freeze_panes = "A4"


def create_leaver_reconciliation_sheet(ws, styles):
    """Create the Leaver_Reconciliation sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "LEAVER RECONCILIATION (Monthly)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Review_Date", 14),
        ("B", "HR_Leavers", 14),
        ("C", "Accounts_Disabled", 18),
        ("D", "Discrepancies", 14),
        ("E", "Orphaned_Accounts", 30),
        ("F", "Remediation_Action", 40),
        ("G", "Remediation_Date", 16),
        ("H", "Reviewer", 25),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Add formula for discrepancy calculation
    for row in range(4, 28):
        for col in ["A", "B", "C", "E", "F", "G", "H"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

        # Discrepancy formula
        ws[f"D{row}"] = f"=IF(B{row}<>\"\",B{row}-C{row},\"\")"
        ws[f"D{row}"].border = styles["border"]
        ws[f"D{row}"].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.freeze_panes = "A4"


def create_evidence_register_sheet(ws, styles):
    """Create the Evidence_Register sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Evidence_ID", 22),
        ("B", "Evidence_Description", 50),
        ("C", "Evidence_Type", 20),
        ("D", "Storage_Location", 50),
        ("E", "Collection_Date", 14),
        ("F", "Collected_By", 25),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    type_val = DataValidation(
        type="list",
        formula1='"Exit Procedure,Access Revocation Log,Asset Return Form,Exit Checklist,Reconciliation Report,Other"'
    )
    ws.add_data_validation(type_val)

    for row in range(4, 54):
        ws[f"A{row}"] = f"EVD-A.6.4-5.2-{row-3:03d}"
        ws[f"A{row}"].font = Font(color="808080")

        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

        type_val.add(ws[f"C{row}"])

    ws.freeze_panes = "A4"


def create_approval_signoff_sheet(ws, styles):
    """Create the Approval_SignOff sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Exit Types Documented", "6 (pre-populated)"),
        ("Access Types Defined", "10 (pre-populated)"),
        ("Asset Categories Defined", "7 (pre-populated)"),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    reviewers = [
        ("IAM Manager", "", ""),
        ("HR Director", "", ""),
        ("CISO", "", ""),
    ]

    row += 1
    ws[f"A{row}"] = "Role"
    ws[f"B{row}"] = "Name"
    ws[f"C{row}"] = "Signature"
    ws[f"D{row}"] = "Date"
    ws[f"E{row}"] = "Comments"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].border = styles["border"]

    row += 1
    for role, _, _ in reviewers:
        ws[f"A{row}"] = role
        ws[f"A{row}"].border = styles["border"]
        for col in ["B", "C", "D", "E"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 35


def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/8] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/8] Creating Exit_Procedures sheet...")
        create_exit_procedures_sheet(wb["Exit_Procedures"], styles)

        logger.info("[3/8] Creating Access_Revocation sheet...")
        create_access_revocation_sheet(wb["Access_Revocation"], styles)

        logger.info("[4/8] Creating Asset_Recovery sheet...")
        create_asset_recovery_sheet(wb["Asset_Recovery"], styles)

        logger.info("[5/8] Creating Exit_Tracker sheet...")
        create_exit_tracker_sheet(wb["Exit_Tracker"], styles)

        logger.info("[6/8] Creating Leaver_Reconciliation sheet...")
        create_leaver_reconciliation_sheet(wb["Leaver_Reconciliation"], styles)

        logger.info("[7/8] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[8/8] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.6.4-5.2 specification
# =============================================================================
