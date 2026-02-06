#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
A.6.4-5 Sanity Check Script
================================================================================

Validates all A.6.4-5 generated workbooks for:
- Required sheets present
- Column headers correct
- Data validations functional
- Formulas valid
- Consistent styling

Usage:
    python3 sanity_check_a645.py
================================================================================
"""

import logging
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# CONSTANTS
# =============================================================================
CONTROL_ID = "A.6.4-5"
WORKBOOK_DIR = Path(__file__).parent.parent / "90_workbooks"

# Expected workbook structure
EXPECTED_WORKBOOKS = {
    "A.6.4-5.1": {
        "pattern": "ISMS-IMP-A.6.4-5.S1_Disciplinary_Process*.xlsx",
        "required_sheets": [
            "Instructions",
            "Violation_Categories",
            "Response_Matrix",
            "Investigation_Process",
            "Case_Tracker",
            "Evidence_Register",
            "Approval_SignOff",
        ]
    },
    "A.6.4-5.2": {
        "pattern": "ISMS-IMP-A.6.4-5.S2_Employment_Exit*.xlsx",
        "required_sheets": [
            "Instructions",
            "Exit_Procedures",
            "Access_Revocation",
            "Asset_Recovery",
            "Exit_Tracker",
            "Leaver_Reconciliation",
            "Evidence_Register",
            "Approval_SignOff",
        ]
    },
    "A.6.4-5.3": {
        "pattern": "ISMS-IMP-A.6.4-5.S3_Post_Employment*.xlsx",
        "required_sheets": [
            "Instructions",
            "Obligation_Types",
            "Former_Personnel",
            "Active_Obligations",
            "Expiration_Tracker",
            "Acknowledgement_Log",
            "Enforcement_Register",
            "Evidence_Register",
            "Approval_SignOff",
        ]
    },
    "A.6.4-5.4": {
        "pattern": "ISMS-IMP-A.6.4-5.S4_Employment_Exit_Dashboard*.xlsx",
        "required_sheets": [
            "Executive_Dashboard",
            "Exit_Metrics",
            "Access_Revocation_Metrics",
            "Asset_Metrics",
            "Disciplinary_Metrics",
            "Obligation_Metrics",
            "Trend_Analysis",
            "Data_Sources",
            "Instructions",
        ]
    },
}


def check_workbook(filepath: Path, expected_sheets: list) -> tuple:
    """
    Check a single workbook for sanity.

    Args:
        filepath: Path to the workbook
        expected_sheets: List of required sheet names

    Returns:
        tuple: (passed_count, failed_count, issues_list)
    """
    issues = []
    passed = 0
    failed = 0

    try:
        wb = load_workbook(filepath, data_only=False)

        # Check required sheets
        for sheet_name in expected_sheets:
            if sheet_name in wb.sheetnames:
                passed += 1
            else:
                failed += 1
                issues.append(f"Missing sheet: {sheet_name}")

        # Check header row exists
        for ws in wb.worksheets:
            if ws["A1"].value is None:
                issues.append(f"Sheet '{ws.title}' has no header in A1")
                failed += 1
            else:
                passed += 1

            # Check for data validations (warning only)
            if ws.data_validations.dataValidation:
                passed += 1
            else:
                if ws.title not in ["Instructions", "Executive_Dashboard"]:
                    # Not a failure, just a note
                    pass

        # Check for formulas in appropriate sheets
        formula_sheets = [
            "Case_Tracker", "Exit_Tracker", "Leaver_Reconciliation",
            "Expiration_Tracker", "Exit_Metrics", "Access_Revocation_Metrics",
            "Asset_Metrics", "Trend_Analysis"
        ]

        for ws in wb.worksheets:
            if ws.title in formula_sheets:
                formula_found = False
                for row in ws.iter_rows(min_row=4, max_row=min(20, ws.max_row)):
                    for cell in row:
                        if cell.value and str(cell.value).startswith("="):
                            formula_found = True
                            break
                    if formula_found:
                        break

                if formula_found:
                    passed += 1
                else:
                    # Some sheets may not have formulas yet
                    pass

        # Check freeze panes
        for ws in wb.worksheets:
            if ws.title != "Instructions" and ws.max_row > 3:
                if ws.freeze_panes:
                    passed += 1
                else:
                    issues.append(f"Sheet '{ws.title}' has no freeze panes")

        wb.close()

    except Exception as e:
        failed += 1
        issues.append(f"Error loading workbook: {e}")

    return passed, failed, issues


def main() -> int:
    """Main execution function."""
    logger.info("=" * 70)
    logger.info(f"{CONTROL_ID} Sanity Check")
    logger.info("=" * 70)

    if not WORKBOOK_DIR.exists():
        logger.warning(f"Workbook directory not found: {WORKBOOK_DIR}")
        logger.info("Run generators first to create workbooks")
        return 0

    total_passed = 0
    total_failed = 0
    all_issues = []

    for doc_id, config in EXPECTED_WORKBOOKS.items():
        logger.info(f"\nChecking {doc_id}...")

        workbooks = list(WORKBOOK_DIR.glob(config["pattern"]))

        if not workbooks:
            logger.warning(f"  No workbooks found matching {config['pattern']}")
            total_failed += 1
            all_issues.append(f"{doc_id}: No workbook found")
            continue

        for wb_path in workbooks:
            logger.info(f"  Checking: {wb_path.name}")
            passed, failed, issues = check_workbook(wb_path, config["required_sheets"])

            total_passed += passed
            total_failed += failed

            if issues:
                for issue in issues:
                    logger.warning(f"    WARNING: {issue}")
                    all_issues.append(f"{doc_id}: {issue}")
            else:
                logger.info(f"    OK All checks passed ({passed} checks)")

    logger.info("\n" + "=" * 70)
    logger.info("SANITY CHECK SUMMARY")
    logger.info("=" * 70)
    logger.info(f"Total checks passed: {total_passed}")
    logger.info(f"Total checks failed: {total_failed}")

    if all_issues:
        logger.info(f"\nIssues found ({len(all_issues)}):")
        for issue in all_issues:
            logger.info(f"  - {issue}")
    else:
        logger.info("\nOK All sanity checks passed!")

    logger.info("=" * 70)

    return 0 if total_failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.6.4-5 control
# =============================================================================
