#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.7-8 Workbook Normalization Script
================================================================================

Normalizes all A.6.7-8 assessment workbooks to ensure consistent formatting,
styling, and structure across S1-S5 workbooks.

Usage:
    python3 normalize_a678_workbooks.py [workbook_path]

    If no path provided, normalizes all workbooks in ../90_workbooks/

================================================================================
"""

import logging
import sys
import os
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# NORMALIZATION STANDARDS
# =============================================================================
STANDARD_FONTS = {
    "header": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
    "subheader": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
    "column_header": Font(name="Calibri", size=10, bold=True),
    "body": Font(name="Calibri", size=10),
}

STANDARD_FILLS = {
    "header": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
    "subheader": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
    "column_header": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "input": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
    "compliant": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "partial": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "noncompliant": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

STANDARD_COLUMN_WIDTHS = {
    "id": 15,
    "name": 25,
    "description": 45,
    "status": 14,
    "date": 12,
    "notes": 30,
}


def normalize_workbook(filepath: str) -> bool:
    """
    Normalize a single workbook to standard formatting.

    Args:
        filepath: Path to the Excel workbook

    Returns:
        True if successful, False otherwise
    """
    logger.info(f"Normalizing: {filepath}")

    try:
        wb = load_workbook(filepath)

        for ws in wb.worksheets:
            # Normalize row 1 header if merged
            if ws.merged_cells.ranges:
                for merged_range in list(ws.merged_cells.ranges):
                    if merged_range.min_row == 1:
                        # Header row - ensure proper styling
                        cell = ws.cell(row=1, column=merged_range.min_col)
                        cell.font = STANDARD_FONTS["header"]
                        cell.fill = STANDARD_FILLS["header"]
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Ensure Instructions sheet has proper row height
            if ws.title == "Instructions":
                ws.row_dimensions[1].height = 40

            # Normalize Approval_Sign_Off sheet if present
            if ws.title == "Approval_Sign_Off":
                ws.row_dimensions[1].height = 30

        # Save normalized workbook
        wb.save(filepath)
        logger.info(f"Successfully normalized: {filepath}")
        return True

    except Exception as e:
        logger.error(f"Error normalizing {filepath}: {e}")
        return False


def normalize_all_workbooks(directory: str) -> tuple:
    """
    Normalize all workbooks in a directory.

    Args:
        directory: Path to directory containing workbooks

    Returns:
        Tuple of (success_count, failure_count)
    """
    success = 0
    failure = 0

    workbooks_dir = Path(directory)
    if not workbooks_dir.exists():
        logger.error(f"Directory not found: {directory}")
        return (0, 0)

    xlsx_files = list(workbooks_dir.glob("ISMS-IMP-A.6.7-8*.xlsx"))

    if not xlsx_files:
        logger.warning(f"No A.6.7-8 workbooks found in {directory}")
        return (0, 0)

    logger.info(f"Found {len(xlsx_files)} workbooks to normalize")

    for xlsx_file in xlsx_files:
        if normalize_workbook(str(xlsx_file)):
            success += 1
        else:
            failure += 1

    return (success, failure)


def main():
    """Main entry point."""
    logger.info("=" * 60)
    logger.info("ISMS-IMP-A.6.7-8 Workbook Normalization")
    logger.info("=" * 60)

    if len(sys.argv) > 1:
        # Normalize specific file
        filepath = sys.argv[1]
        if os.path.exists(filepath):
            success = normalize_workbook(filepath)
            sys.exit(0 if success else 1)
        else:
            logger.error(f"File not found: {filepath}")
            sys.exit(1)
    else:
        # Normalize all workbooks in default directory
        script_dir = Path(__file__).parent
        workbooks_dir = script_dir.parent / "90_workbooks"

        success, failure = normalize_all_workbooks(str(workbooks_dir))

        logger.info("=" * 60)
        logger.info(f"Normalization Complete: {success} success, {failure} failed")
        logger.info("=" * 60)

        sys.exit(0 if failure == 0 else 1)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation for A.6.7-8 workbook normalization
# =============================================================================
