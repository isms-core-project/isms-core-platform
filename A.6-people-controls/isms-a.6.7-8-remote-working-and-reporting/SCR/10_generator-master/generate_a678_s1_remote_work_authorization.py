#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.7-8.S1 - Remote Work Authorization Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.6.7 (Remote Working) & A.6.8 (Event Reporting)
Assessment Domain S1 of S5: Remote Work Authorization Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific remote work policies, authorization requirements,
and risk assessment criteria.

Key customization areas:
1. Role definitions and remote work eligibility
2. Authorization approval workflows
3. Risk assessment criteria
4. Physical security self-assessment questions
5. Regulatory requirements specific to your jurisdiction

Reference Pattern: Based on ISMS-IMP-A.6.7-8.S1 specification

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates an Excel workbook for conducting Remote Work Authorization
Assessment per ISO 27001:2022 Controls A.6.7 and A.6.8 requirements.

**Purpose:**
Enables systematic assessment of remote work authorization processes including
eligibility determination, risk assessment, and approval workflows.

**Assessment Scope:**
- Remote work eligibility criteria
- Authorization request and approval processes
- Risk assessment for remote roles
- Physical security self-assessment
- Policy acknowledgment tracking

**Generated Workbook Structure:**
1. Instructions - Assessment guidance
2. Eligibility_Criteria - Role eligibility assessment
3. Authorization_Register - Remote work authorizations
4. Risk_Assessment - Role-based risk evaluation
5. Physical_Security - Workspace self-assessment
6. Acknowledgments - Policy acknowledgment tracking
7. Gap_Analysis - Identified gaps
8. Evidence_Register - Audit evidence
9. Dashboard - Summary metrics
10. Approval_Sign_Off - Stakeholder approval

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    pip3 install openpyxl

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.6.7, A.6.8
Assessment Domain:    S1 of S5 (Remote Work Authorization)
Framework Version:    1.0
Script Version:       1.0
Author:               ISMS Core Contributors
Created:              2026
Last Modified:        [Date to be set]

Related Documents:
    - ISMS-POL-A.6.7-8: Remote Working and Security Event Reporting Policy
    - ISMS-IMP-A.6.7-8.S2: Technical Controls Assessment
    - ISMS-IMP-A.6.7-8.S3: Endpoint and Physical Security Assessment
    - ISMS-IMP-A.6.7-8.S4: Event Reporting Mechanisms Assessment
    - ISMS-IMP-A.6.7-8.S5: Compliance Dashboard

================================================================================
"""

# =============================================================================
# IMPORTS - Standard Library
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# IMPORTS - Third Party
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.7-8.S1"
WORKBOOK_NAME = "Remote Work Authorization Assessment"
CONTROL_ID = "A.6.7-8"
CONTROL_NAME = "Remote Working and Security Event Reporting"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # Swiss format
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# =============================================================================
# PRE-POPULATED DATA
# =============================================================================

# Eligibility Criteria per POL-A.6.7-8 Section 2.1
ELIGIBILITY_CRITERIA = [
    ("Role Suitability", "Role can be performed remotely without security compromise", "High"),
    ("Data Classification", "Data accessed remotely is classified and handling requirements defined", "High"),
    ("Technical Capability", "Personnel can establish secure remote connections", "High"),
    ("Physical Environment", "Remote workspace meets privacy and security requirements", "Medium"),
    ("Regulatory Restrictions", "No regulatory or contractual restrictions on remote work", "High"),
    ("Business Continuity", "Remote work supports business continuity requirements", "Medium"),
]

# Risk Assessment Criteria per POL-A.6.7-8 Section 2.1.1
RISK_CRITERIA = [
    ("Data Classification Level", "Public=1, Internal=2, Confidential=3, Restricted=4", 4),
    ("Physical Security Capability", "Excellent=1, Good=2, Adequate=3, Poor=4", 4),
    ("Network Security Posture", "Corporate VPN=1, Secure Home=2, Variable=3, Risky=4", 4),
    ("Device Security Baseline", "Managed Corporate=1, BYOD Compliant=2, Partial=3, Non-compliant=4", 4),
    ("Regulatory Restrictions", "None=1, General=2, Industry-Specific=3, Multiple=4", 4),
]

# Physical Security Self-Assessment Items
PHYSICAL_SECURITY_ITEMS = [
    ("Screen Positioning", "Screen positioned away from windows and common areas", "Mandatory"),
    ("Privacy Screen", "Privacy screen filter available for shared/public spaces", "Conditional"),
    ("Equipment Security", "Workspace has secure storage for devices when unattended", "Mandatory"),
    ("Access Control", "Workspace prevents unauthorized access by family/visitors", "Mandatory"),
    ("Document Storage", "Secure storage available for sensitive documents", "Mandatory"),
    ("Document Disposal", "Access to shredder or secure disposal method", "Mandatory"),
    ("Clear Desk", "Clear desk practice maintained at end of work sessions", "Mandatory"),
]

# Authorization Status Options
AUTHORIZATION_STATUS = ["Pending", "Approved", "Conditionally Approved", "Denied", "Revoked", "Expired"]


# =============================================================================
# WORKBOOK CREATION
# =============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Eligibility_Criteria",
        "Authorization_Register",
        "Risk_Assessment",
        "Physical_Security",
        "Acknowledgments",
        "Gap_Analysis",
        "Evidence_Register",
        "Dashboard",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# SHEET: INSTRUCTIONS
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Remote Work Authorization Framework"),
        ("Related Policy", "ISMS-POL-A.6.7-8, Section 2.1 (Remote Work Authorization)"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organization", ""),
        ("Review Cycle", "Annual + upon significant changes to remote work arrangements"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Review Eligibility_Criteria to define which roles are eligible for remote work.",
        "2. Complete Authorization_Register for all current and pending remote work arrangements.",
        "3. Use Risk_Assessment to evaluate security risks for each remote work arrangement.",
        "4. Ensure Physical_Security self-assessments are completed by remote workers.",
        "5. Track policy Acknowledgments for all remote workers.",
        "6. Document any gaps identified in Gap_Analysis.",
        "7. Maintain Evidence_Register for audit traceability.",
        "8. Review Dashboard for summary metrics.",
        "9. Obtain approvals in Approval_Sign_Off sheet.",
    ]

    row += 1
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    row += 1
    ws[f"A{row}"] = "AUTHORIZATION WORKFLOW"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    workflow = [
        "Step 1: Employee requests remote work arrangement",
        "Step 2: Line Manager evaluates eligibility and role suitability",
        "Step 3: IT Security reviews technical requirements and performs risk assessment",
        "Step 4: Employee completes physical security self-assessment",
        "Step 5: Employee acknowledges remote work security requirements",
        "Step 6: Authorization granted or conditionally approved with controls",
        "Step 7: Annual review of all active authorizations",
    ]

    row += 1
    for line in workflow:
        ws[f"A{row}"] = line
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50


# =============================================================================
# SHEET: ELIGIBILITY CRITERIA
# =============================================================================

def create_eligibility_sheet(ws, styles):
    """Create the Eligibility Criteria sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "Remote Work Eligibility Criteria Assessment"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Criterion", "Description", "Importance", "Documented", "Evidence", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row = 4
    for criterion, description, importance in ELIGIBILITY_CRITERIA:
        ws.cell(row=row, column=1, value=criterion).border = styles["border"]
        ws.cell(row=row, column=2, value=description).border = styles["border"]
        ws.cell(row=row, column=3, value=importance).border = styles["border"]
        # Input cells
        for col in range(4, 7):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Add data validation for Documented column
    dv = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"D4:D{row-1}")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 30


# =============================================================================
# SHEET: AUTHORIZATION REGISTER
# =============================================================================

def create_authorization_register_sheet(ws, styles):
    """Create the Authorization Register sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "Remote Work Authorization Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = [
        "Auth ID", "Employee Name", "Department", "Role", "Remote Type",
        "Request Date", "Risk Assessment", "Physical Assessment",
        "Acknowledgment", "Status", "Approval Date", "Review Date"
    ]

    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Add 20 empty rows for data entry
    for data_row in range(4, 24):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Add data validations
    remote_type_dv = DataValidation(type="list", formula1='"Full-Time,Part-Time,Occasional,Travel Only"', allow_blank=True)
    ws.add_data_validation(remote_type_dv)
    remote_type_dv.add("E4:E23")

    status_dv = DataValidation(type="list", formula1=f'"{",".join(AUTHORIZATION_STATUS)}"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("J4:J23")

    completion_dv = DataValidation(type="list", formula1='"Complete,Pending,N/A"', allow_blank=True)
    ws.add_data_validation(completion_dv)
    completion_dv.add("G4:I23")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


# =============================================================================
# SHEET: RISK ASSESSMENT
# =============================================================================

def create_risk_assessment_sheet(ws, styles):
    """Create the Risk Assessment sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Remote Work Risk Assessment"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Risk Criteria Reference
    ws["A3"] = "Risk Scoring Criteria"
    ws["A3"].font = Font(bold=True, size=11)

    headers = ["Criterion", "Scoring Guide", "Max Score"]
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row = 5
    for criterion, scoring, max_score in RISK_CRITERIA:
        ws.cell(row=row, column=1, value=criterion).border = styles["border"]
        ws.cell(row=row, column=2, value=scoring).border = styles["border"]
        ws.cell(row=row, column=3, value=max_score).border = styles["border"]
        row += 1

    row += 2
    ws[f"A{row}"] = "Risk Assessment Register"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    assess_headers = ["Auth ID", "Employee", "Data Class", "Physical", "Network", "Device", "Regulatory", "Total", "Risk Level"]
    for col_idx, header in enumerate(assess_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Add 15 empty rows for assessments
    for data_row in range(row + 1, row + 16):
        for col in range(1, len(assess_headers) + 1):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Risk level validation
    risk_dv = DataValidation(type="list", formula1='"Low,Medium,High,Critical"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add(f"I{row+1}:I{row+15}")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 12


# =============================================================================
# SHEET: PHYSICAL SECURITY
# =============================================================================

def create_physical_security_sheet(ws, styles):
    """Create the Physical Security Self-Assessment sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "Physical Security Self-Assessment Template"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Requirements reference
    ws["A3"] = "Physical Security Requirements"
    ws["A3"].font = Font(bold=True, size=11)

    headers = ["Requirement", "Description", "Requirement Level", "Verification Method"]
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row = 5
    for requirement, description, level in PHYSICAL_SECURITY_ITEMS:
        ws.cell(row=row, column=1, value=requirement).border = styles["border"]
        ws.cell(row=row, column=2, value=description).border = styles["border"]
        ws.cell(row=row, column=3, value=level).border = styles["border"]
        ws.cell(row=row, column=4, value="Self-assessment + Annual attestation").border = styles["border"]
        row += 1

    row += 2
    ws[f"A{row}"] = "Individual Self-Assessments"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    assess_headers = ["Employee", "Assessment Date", "All Mandatory Met", "Conditional Items", "Overall Status", "Next Review", "Notes"]
    for col_idx, header in enumerate(assess_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Add 15 empty rows
    for data_row in range(row + 1, row + 16):
        for col in range(1, len(assess_headers) + 1):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Validations
    yes_no_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    yes_no_dv.add(f"C{row+1}:C{row+15}")

    status_dv = DataValidation(type="list", formula1='"Compliant,Non-Compliant,Pending Review"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f"E{row+1}:E{row+15}")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 30


# =============================================================================
# SHEET: ACKNOWLEDGMENTS
# =============================================================================

def create_acknowledgments_sheet(ws, styles):
    """Create the Policy Acknowledgments tracking sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Remote Work Policy Acknowledgment Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = [
        "Employee Name", "Department", "Policy Version", "Acknowledgment Date",
        "Method", "Expiry Date", "Status", "Evidence Reference"
    ]

    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Add 20 empty rows
    for data_row in range(4, 24):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Validations
    method_dv = DataValidation(type="list", formula1='"Electronic Signature,Physical Signature,LMS Completion,Email Confirmation"', allow_blank=True)
    ws.add_data_validation(method_dv)
    method_dv.add("E4:E23")

    status_dv = DataValidation(type="list", formula1='"Current,Expired,Pending"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("G4:G23")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


# =============================================================================
# SHEET: GAP ANALYSIS
# =============================================================================

def create_gap_analysis_sheet(ws, styles):
    """Create the Gap Analysis sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Remote Work Authorization Gap Analysis"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = [
        "Gap ID", "Source Area", "Gap Description", "Affected Scope",
        "Risk Level", "Remediation Action", "Owner", "Target Date", "Status"
    ]

    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Add 15 empty rows with auto-generated IDs
    for data_row in range(4, 19):
        ws.cell(row=data_row, column=1, value=f"GAP-RWA-{data_row-3:03d}").border = styles["border"]
        for col in range(2, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Validations
    source_dv = DataValidation(type="list", formula1='"Eligibility,Authorization,Risk Assessment,Physical Security,Acknowledgment"', allow_blank=True)
    ws.add_data_validation(source_dv)
    source_dv.add("B4:B18")

    risk_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add("E4:E18")

    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Accepted"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("I4:I18")

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14


# =============================================================================
# SHEET: EVIDENCE REGISTER
# =============================================================================

def create_evidence_register_sheet(ws, styles):
    """Create the Evidence Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Remote Work Authorization Evidence Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = [
        "Evidence ID", "Evidence Type", "Description", "Source",
        "Date Collected", "Retention Period", "Location", "Status"
    ]

    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Pre-populate evidence types
    evidence_types = [
        ("EVD-RWA-001", "Authorization Forms", "Completed remote work authorization requests", "HR/IT"),
        ("EVD-RWA-002", "Risk Assessments", "Documented risk assessments for remote roles", "IT Security"),
        ("EVD-RWA-003", "Physical Assessments", "Physical security self-assessment forms", "Remote Workers"),
        ("EVD-RWA-004", "Policy Acknowledgments", "Signed policy acknowledgment records", "HR"),
        ("EVD-RWA-005", "Approval Records", "Authorization approval documentation", "Line Managers"),
        ("EVD-RWA-006", "Review Records", "Annual authorization review records", "IT Security"),
    ]

    row = 4
    for evd_id, evd_type, description, source in evidence_types:
        ws.cell(row=row, column=1, value=evd_id).border = styles["border"]
        ws.cell(row=row, column=2, value=evd_type).border = styles["border"]
        ws.cell(row=row, column=3, value=description).border = styles["border"]
        ws.cell(row=row, column=4, value=source).border = styles["border"]
        for col in range(5, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Status validation
    status_dv = DataValidation(type="list", formula1='"Collected,Pending,Not Available"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f"H4:H{row-1}")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


# =============================================================================
# SHEET: DASHBOARD
# =============================================================================

def create_dashboard_sheet(ws, styles):
    """Create the Dashboard sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "Remote Work Authorization Dashboard"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Summary Metrics
    ws["A3"] = "Summary Metrics"
    ws["A3"].font = Font(bold=True, size=12)

    metrics = [
        ("Total Remote Work Authorizations", "=COUNTA(Authorization_Register!A4:A23)"),
        ("Approved Authorizations", '=COUNTIF(Authorization_Register!J4:J23,"Approved")'),
        ("Pending Authorizations", '=COUNTIF(Authorization_Register!J4:J23,"Pending")'),
        ("Revoked/Expired", '=COUNTIF(Authorization_Register!J4:J23,"Revoked")+COUNTIF(Authorization_Register!J4:J23,"Expired")'),
        ("Physical Assessments Complete", '=COUNTIF(Physical_Security!E14:E28,"Compliant")'),
        ("Policy Acknowledgments Current", '=COUNTIF(Acknowledgments!G4:G23,"Current")'),
        ("Open Gaps", '=COUNTIF(Gap_Analysis!I4:I18,"Open")'),
        ("High/Critical Risk Items", '=COUNTIF(Risk_Assessment!I14:I28,"High")+COUNTIF(Risk_Assessment!I14:I28,"Critical")'),
    ]

    row = 4
    for metric, formula in metrics:
        ws[f"A{row}"] = metric
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = formula
        ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws[f"A{row}"] = "Compliance Status"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    compliance_items = [
        ("Authorization Process Documented", ""),
        ("Risk Assessments Performed", ""),
        ("Physical Security Verified", ""),
        ("All Acknowledgments Current", ""),
        ("Annual Reviews Scheduled", ""),
    ]

    for item, status in compliance_items:
        ws[f"A{row}"] = item
        ws[f"A{row}"].border = styles["border"]
        cell = ws[f"B{row}"]
        cell.fill = styles["input_cell"]["fill"]
        cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20


# =============================================================================
# SHEET: APPROVAL SIGN-OFF
# =============================================================================

def create_approval_sheet(ws, styles):
    """Create the Approval Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "Assessment Approval and Sign-Off"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Role", "Name", "Signature", "Date", "Comments"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    approvers = [
        "IT Security Manager",
        "HR Manager",
        "Line Manager Representative",
        "CISO",
    ]

    row = 4
    for approver in approvers:
        ws.cell(row=row, column=1, value=approver).border = styles["border"]
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 40


# =============================================================================
# MAIN FUNCTION
# =============================================================================

def main():
    """Generate the Remote Work Authorization Assessment workbook."""
    logger.info(f"Starting generation of {DOCUMENT_ID} - {WORKBOOK_NAME}")
    logger.info(f"Output file: {OUTPUT_FILENAME}")

    try:
        # Setup
        styles = setup_styles()
        wb = create_workbook()

        # Create sheets
        logger.info("Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("Creating Eligibility Criteria sheet...")
        create_eligibility_sheet(wb["Eligibility_Criteria"], styles)

        logger.info("Creating Authorization Register sheet...")
        create_authorization_register_sheet(wb["Authorization_Register"], styles)

        logger.info("Creating Risk Assessment sheet...")
        create_risk_assessment_sheet(wb["Risk_Assessment"], styles)

        logger.info("Creating Physical Security sheet...")
        create_physical_security_sheet(wb["Physical_Security"], styles)

        logger.info("Creating Acknowledgments sheet...")
        create_acknowledgments_sheet(wb["Acknowledgments"], styles)

        logger.info("Creating Gap Analysis sheet...")
        create_gap_analysis_sheet(wb["Gap_Analysis"], styles)

        logger.info("Creating Evidence Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("Creating Dashboard sheet...")
        create_dashboard_sheet(wb["Dashboard"], styles)

        logger.info("Creating Approval Sign-Off sheet...")
        create_approval_sheet(wb["Approval_Sign_Off"], styles)

        # Save workbook
        wb.save(OUTPUT_FILENAME)
        logger.info(f"Workbook saved successfully: {OUTPUT_FILENAME}")

    except Exception as e:
        logger.error(f"Error generating workbook: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()


# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation, constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
