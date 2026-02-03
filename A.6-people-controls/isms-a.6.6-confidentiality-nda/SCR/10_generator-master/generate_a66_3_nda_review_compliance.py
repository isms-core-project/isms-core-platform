#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.6.3 - NDA Review & Compliance Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.6: Confidentiality or Non-Disclosure Agreements

Assessment Domain 3 of 4: NDA Review & Compliance

This script generates a workbook for periodic review of NDA adequacy,
compliance verification, and gap identification.

**Generated Workbook Structure:**
1. Instructions - Review guidance
2. Periodic_Review - Scheduled review tracking
3. Template_Adequacy - Template sufficiency assessment
4. Coverage_Analysis - NDA coverage verification
5. Compliance_Check - Signatory compliance verification
6. Gap_Register - Identified gaps
7. Evidence_Register - Audit evidence tracking
8. Approval_SignOff - Stakeholder approval

================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

DOCUMENT_ID = "ISMS-IMP-A.6.6.3"
WORKBOOK_NAME = "NDA Review and Compliance"
CONTROL_ID = "A.6.6"
CONTROL_NAME = "Confidentiality or Non-Disclosure Agreements"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


def setup_styles():
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }


def create_workbook() -> Workbook:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    for name in ["Instructions", "Periodic_Review", "Template_Adequacy",
                 "Coverage_Analysis", "Compliance_Check", "Gap_Register",
                 "Evidence_Register", "Approval_SignOff"]:
        wb.create_sheet(title=name)
    return wb


def create_instructions_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:H3")
    ws["A3"] = CONTROL_REF
    ws["A3"].font = Font(bold=True, size=12)

    instructions = [
        "",
        "PURPOSE",
        "This workbook enables periodic review of NDA adequacy and compliance,",
        "ensuring NDAs remain current, appropriate, and properly executed.",
        "",
        "ISO 27001:2022 REQUIREMENT",
        "NDAs should be 'regularly reviewed' - this workbook tracks that requirement.",
        "",
        "SCOPE",
        "- Periodic review scheduling and tracking",
        "- Template adequacy assessment",
        "- Coverage analysis (who should have NDAs vs. who does)",
        "- Compliance verification (signed, current, appropriate)",
        "- Gap identification and remediation",
        "",
        "COMPLETION STEPS",
        "1. Periodic_Review: Schedule and track review activities",
        "2. Template_Adequacy: Assess template sufficiency",
        "3. Coverage_Analysis: Verify all required persons have NDAs",
        "4. Compliance_Check: Verify NDA currency and appropriateness",
        "5. Gap_Register: Document identified gaps",
        "6. Evidence_Register: Link supporting evidence",
        "7. Approval_SignOff: Obtain required approvals",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, text in enumerate(instructions, start=5):
        ws[f"A{i}"] = text
        if text in ["PURPOSE", "ISO 27001:2022 REQUIREMENT", "SCOPE", "COMPLETION STEPS"]:
            ws[f"A{i}"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 80


def create_periodic_review_sheet(ws, styles):
    ws.merge_cells("A1:K1")
    ws["A1"] = "Periodic Review Schedule & Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Review_ID", 14),
        ("B", "Review_Type", 20),
        ("C", "Review_Scope", 30),
        ("D", "Planned_Date", 14),
        ("E", "Actual_Date", 14),
        ("F", "Reviewer", 22),
        ("G", "Findings_Count", 14),
        ("H", "Gaps_Identified", 14),
        ("I", "Status", 14),
        ("J", "Next_Review", 14),
        ("K", "Notes", 35),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    type_val = DataValidation(type="list", formula1='"Annual Full Review,Quarterly Check,Template Update Review,Triggered Review,Ad-hoc Review"')
    type_val.add("B3:B50")
    ws.add_data_validation(type_val)

    status_val = DataValidation(type="list", formula1='"Scheduled,In Progress,Completed,Overdue,Cancelled"')
    status_val.add("I3:I50")
    ws.add_data_validation(status_val)

    for row in range(3, 15):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_template_adequacy_sheet(ws, styles):
    ws.merge_cells("A1:L1")
    ws["A1"] = "NDA Template Adequacy Assessment"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Template_ID", 14),
        ("B", "Template_Name", 30),
        ("C", "Last_Legal_Review", 16),
        ("D", "Regulatory_Current", 16),
        ("E", "Covers_All_Info_Types", 18),
        ("F", "Post_Term_Adequate", 16),
        ("G", "Remedies_Adequate", 16),
        ("H", "Jurisdiction_Correct", 16),
        ("I", "Overall_Adequacy", 16),
        ("J", "Score", 10),
        ("K", "Action_Required", 30),
        ("L", "Notes", 30),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    adequacy_val = DataValidation(type="list", formula1='"Adequate,Partially Adequate,Inadequate,Not Assessed"')
    adequacy_val.add("D3:D50")
    adequacy_val.add("E3:E50")
    adequacy_val.add("F3:F50")
    adequacy_val.add("G3:G50")
    adequacy_val.add("H3:H50")
    adequacy_val.add("I3:I50")
    ws.add_data_validation(adequacy_val)

    for row in range(3, 15):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_coverage_analysis_sheet(ws, styles):
    ws.merge_cells("A1:J1")
    ws["A1"] = "NDA Coverage Analysis"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Stakeholder_Category", 22),
        ("B", "Total_Count", 12),
        ("C", "NDA_Required", 12),
        ("D", "NDA_Signed", 12),
        ("E", "Coverage_Rate", 14),
        ("F", "Expired_NDAs", 12),
        ("G", "Missing_NDAs", 12),
        ("H", "Gap_Status", 14),
        ("I", "Remediation_Owner", 20),
        ("J", "Notes", 35),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Pre-populate categories
    categories = [
        "Employees",
        "Contractors",
        "Consultants",
        "Vendors",
        "Suppliers",
        "Partners",
        "Customers",
        "Board Members",
        "Visitors",
        "Other",
    ]

    for row_idx, category in enumerate(categories, start=3):
        ws[f"A{row_idx}"] = category
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            if col != "A":
                ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

    gap_val = DataValidation(type="list", formula1='"No Gap,Gap Identified,Remediation In Progress,Remediated"')
    gap_val.add("H3:H20")
    ws.add_data_validation(gap_val)

    ws.freeze_panes = "A3"


def create_compliance_check_sheet(ws, styles):
    ws.merge_cells("A1:K1")
    ws["A1"] = "NDA Compliance Verification"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "NDA_ID", 14),
        ("B", "Counterparty", 25),
        ("C", "Correctly_Executed", 16),
        ("D", "Within_Validity", 14),
        ("E", "Appropriate_Template", 18),
        ("F", "All_Parties_Signed", 16),
        ("G", "Securely_Stored", 14),
        ("H", "Overall_Compliance", 16),
        ("I", "Issues_Found", 35),
        ("J", "Action_Required", 30),
        ("K", "Status", 14),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    yes_no = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"')
    yes_no.add("C3:C100")
    yes_no.add("D3:D100")
    yes_no.add("E3:E100")
    yes_no.add("F3:F100")
    yes_no.add("G3:G100")
    ws.add_data_validation(yes_no)

    compliance_val = DataValidation(type="list", formula1='"Compliant,Partially Compliant,Non-Compliant"')
    compliance_val.add("H3:H100")
    ws.add_data_validation(compliance_val)

    status_val = DataValidation(type="list", formula1='"No Action Needed,Action Required,In Progress,Resolved"')
    status_val.add("K3:K100")
    ws.add_data_validation(status_val)

    for row in range(3, 23):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_gap_register_sheet(ws, styles):
    ws.merge_cells("A1:L1")
    ws["A1"] = "NDA Gap Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Gap_ID", 12),
        ("B", "Gap_Type", 18),
        ("C", "Description", 45),
        ("D", "Affected_Area", 22),
        ("E", "Severity", 12),
        ("F", "Identified_Date", 14),
        ("G", "Owner", 20),
        ("H", "Remediation_Action", 40),
        ("I", "Target_Date", 14),
        ("J", "Status", 14),
        ("K", "Closure_Date", 14),
        ("L", "Notes", 30),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    type_val = DataValidation(type="list", formula1='"Missing NDA,Expired NDA,Inadequate Template,Unsigned,Wrong Template,Storage Issue,Other"')
    type_val.add("B3:B50")
    ws.add_data_validation(type_val)

    severity_val = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    severity_val.add("E3:E50")
    ws.add_data_validation(severity_val)

    status_val = DataValidation(type="list", formula1='"Open,In Progress,Remediated,Verified Closed,Risk Accepted"')
    status_val.add("J3:J50")
    ws.add_data_validation(status_val)

    for row in range(3, 23):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_evidence_register_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "Evidence Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Evidence_ID", 14),
        ("B", "Review_Ref", 14),
        ("C", "Evidence_Type", 22),
        ("D", "Description", 40),
        ("E", "Storage_Location", 35),
        ("F", "Collected_Date", 14),
        ("G", "Collected_By", 20),
        ("H", "Retention_Until", 14),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    type_val = DataValidation(type="list", formula1='"Review Report,Coverage Analysis,Compliance Check,Gap Evidence,Remediation Evidence,Other"')
    type_val.add("C3:C50")
    ws.add_data_validation(type_val)

    for row in range(3, 23):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_approval_sheet(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "Approval and Sign-Off"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Approval_Type", 25),
        ("B", "Approver_Role", 25),
        ("C", "Approver_Name", 25),
        ("D", "Signature", 20),
        ("E", "Date", 14),
        ("F", "Comments", 35),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    approvals = [
        ("Periodic Review Complete", "Information Security Manager"),
        ("Coverage Analysis Verified", "HR Manager / Contracts Manager"),
        ("Gap Remediation Verified", "CISO"),
        ("Overall Approval", "CISO"),
    ]

    for row_idx, (approval_type, role) in enumerate(approvals, start=3):
        ws[f"A{row_idx}"] = approval_type
        ws[f"B{row_idx}"] = role
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            if col in ["C", "D", "E", "F"]:
                ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def main() -> int:
    logger.info("=" * 60)
    logger.info("Generating %s", OUTPUT_FILENAME)
    logger.info("=" * 60)

    try:
        wb = create_workbook()
        styles = setup_styles()

        create_instructions_sheet(wb["Instructions"], styles)
        create_periodic_review_sheet(wb["Periodic_Review"], styles)
        create_template_adequacy_sheet(wb["Template_Adequacy"], styles)
        create_coverage_analysis_sheet(wb["Coverage_Analysis"], styles)
        create_compliance_check_sheet(wb["Compliance_Check"], styles)
        create_gap_register_sheet(wb["Gap_Register"], styles)
        create_evidence_register_sheet(wb["Evidence_Register"], styles)
        create_approval_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)
        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        return 0

    except Exception as e:
        logger.error("FAILED: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# =============================================================================
