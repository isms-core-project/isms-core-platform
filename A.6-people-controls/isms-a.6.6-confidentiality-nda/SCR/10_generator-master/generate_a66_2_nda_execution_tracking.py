#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.6.2 - NDA Execution & Tracking Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.6: Confidentiality or Non-Disclosure Agreements

Assessment Domain 2 of 4: NDA Execution & Tracking

This script generates a workbook for tracking signed NDAs, signatories,
execution dates, expiration monitoring, and renewal management.

**Generated Workbook Structure:**
1. Instructions - Execution tracking guidance
2. Active_NDAs - Currently active signed NDAs
3. Signatory_Register - Individual signatory tracking
4. Expiration_Monitor - Upcoming expirations
5. Renewal_Tracking - Renewal management
6. Evidence_Register - Audit evidence tracking
7. Approval_SignOff - Stakeholder approval

================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

DOCUMENT_ID = "ISMS-IMP-A.6.6.2"
WORKBOOK_NAME = "NDA Execution and Tracking"
CONTROL_ID = "A.6.6"
CONTROL_NAME = "Confidentiality or Non-Disclosure Agreements"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


def setup_styles():
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "warning": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "alert": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }


def create_workbook() -> Workbook:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    for name in ["Instructions", "Active_NDAs", "Signatory_Register",
                 "Expiration_Monitor", "Renewal_Tracking",
                 "Evidence_Register", "Approval_SignOff"]:
        wb.create_sheet(title=name)
    return wb


def create_instructions_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:H3")
    ws["A3"] = CONTROL_REF
    ws["A3"].font = Font(bold=True, size=12)

    instructions = [
        "",
        "PURPOSE",
        "This workbook tracks all executed (signed) NDAs, individual signatories,",
        "expiration dates, and renewal requirements.",
        "",
        "SCOPE",
        "- All signed confidentiality and non-disclosure agreements",
        "- Individual signatory tracking by stakeholder category",
        "- Expiration monitoring and alerts",
        "- Renewal management and tracking",
        "",
        "COMPLETION STEPS",
        "1. Active_NDAs: Register all currently active signed NDAs",
        "2. Signatory_Register: Track individual signatories",
        "3. Expiration_Monitor: Monitor upcoming expirations",
        "4. Renewal_Tracking: Manage renewal process",
        "5. Evidence_Register: Link supporting evidence",
        "6. Approval_SignOff: Obtain required approvals",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, text in enumerate(instructions, start=5):
        ws[f"A{i}"] = text
        if text in ["PURPOSE", "SCOPE", "COMPLETION STEPS"]:
            ws[f"A{i}"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 80


def create_active_ndas_sheet(ws, styles):
    ws.merge_cells("A1:N1")
    ws["A1"] = "Active NDA Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "NDA_ID", 14),
        ("B", "Template_Ref", 14),
        ("C", "NDA_Title", 35),
        ("D", "Counterparty", 30),
        ("E", "Counterparty_Type", 18),
        ("F", "Execution_Date", 14),
        ("G", "Effective_Date", 14),
        ("H", "Expiration_Date", 14),
        ("I", "Post_Term_Period", 16),
        ("J", "Post_Term_Expiry", 16),
        ("K", "Signatories_Count", 14),
        ("L", "Storage_Location", 30),
        ("M", "Status", 14),
        ("N", "Notes", 30),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    type_val = DataValidation(type="list", formula1='"Employee,Contractor,Consultant,Vendor,Supplier,Partner,Customer,Board Member,Visitor,Other"')
    type_val.add("E3:E200")
    ws.add_data_validation(type_val)

    status_val = DataValidation(type="list", formula1='"Active,Expired,Terminated,Renewed,Superseded"')
    status_val.add("M3:M200")
    ws.add_data_validation(status_val)

    for row in range(3, 23):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_signatory_register_sheet(ws, styles):
    ws.merge_cells("A1:L1")
    ws["A1"] = "Individual Signatory Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Signatory_ID", 14),
        ("B", "NDA_Ref", 14),
        ("C", "Signatory_Name", 25),
        ("D", "Signatory_Type", 18),
        ("E", "Organisation", 25),
        ("F", "Role_Title", 25),
        ("G", "Email", 30),
        ("H", "Signature_Date", 14),
        ("I", "Signature_Method", 16),
        ("J", "Termination_Date", 14),
        ("K", "Status", 14),
        ("L", "Notes", 30),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    type_val = DataValidation(type="list", formula1='"Employee,Contractor,Consultant,Vendor Rep,Partner Rep,Customer Rep,Board Member,Visitor,Witness,Authorised Signatory"')
    type_val.add("D3:D500")
    ws.add_data_validation(type_val)

    method_val = DataValidation(type="list", formula1='"Wet Signature,Digital Signature,Electronic Signature,DocuSign,Adobe Sign,Other"')
    method_val.add("I3:I500")
    ws.add_data_validation(method_val)

    status_val = DataValidation(type="list", formula1='"Active,Terminated,Renewed,Superseded"')
    status_val.add("K3:K500")
    ws.add_data_validation(status_val)

    for row in range(3, 33):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_expiration_monitor_sheet(ws, styles):
    ws.merge_cells("A1:J1")
    ws["A1"] = "NDA Expiration Monitor"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "NDA_ID", 14),
        ("B", "Counterparty", 30),
        ("C", "Expiration_Date", 14),
        ("D", "Days_Until_Expiry", 16),
        ("E", "Alert_Status", 14),
        ("F", "Renewal_Required", 14),
        ("G", "Renewal_Owner", 20),
        ("H", "Renewal_Started", 14),
        ("I", "Action_Required", 30),
        ("J", "Notes", 30),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    alert_val = DataValidation(type="list", formula1='"Green (>90 days),Amber (30-90 days),Red (<30 days),Expired"')
    alert_val.add("E3:E100")
    ws.add_data_validation(alert_val)

    yes_no = DataValidation(type="list", formula1='"Yes,No,Under Review"')
    yes_no.add("F3:F100")
    ws.add_data_validation(yes_no)

    for row in range(3, 23):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_renewal_tracking_sheet(ws, styles):
    ws.merge_cells("A1:L1")
    ws["A1"] = "NDA Renewal Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Renewal_ID", 14),
        ("B", "Original_NDA", 14),
        ("C", "Counterparty", 25),
        ("D", "Original_Expiry", 14),
        ("E", "Renewal_Initiated", 14),
        ("F", "New_Terms_Required", 16),
        ("G", "Legal_Review", 14),
        ("H", "Counterparty_Agreed", 16),
        ("I", "New_NDA_ID", 14),
        ("J", "New_Expiry", 14),
        ("K", "Status", 14),
        ("L", "Notes", 30),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    yes_no = DataValidation(type="list", formula1='"Yes,No,Pending"')
    yes_no.add("F3:F100")
    yes_no.add("G3:G100")
    yes_no.add("H3:H100")
    ws.add_data_validation(yes_no)

    status_val = DataValidation(type="list", formula1='"Not Started,In Progress,Legal Review,Awaiting Signature,Completed,Cancelled"')
    status_val.add("K3:K100")
    ws.add_data_validation(status_val)

    for row in range(3, 23):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_evidence_register_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "Evidence Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Evidence_ID", 14),
        ("B", "NDA_Ref", 14),
        ("C", "Evidence_Type", 22),
        ("D", "Description", 40),
        ("E", "Storage_Location", 35),
        ("F", "Collected_Date", 14),
        ("G", "Collected_By", 20),
        ("H", "Retention_Until", 14),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    type_val = DataValidation(type="list", formula1='"Signed NDA,Signature Page,Email Confirmation,DocuSign Certificate,Renewal Documentation,Termination Notice,Other"')
    type_val.add("C3:C100")
    ws.add_data_validation(type_val)

    for row in range(3, 23):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_approval_sheet(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "Approval and Sign-Off"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Approval_Type", 25),
        ("B", "Approver_Role", 25),
        ("C", "Approver_Name", 25),
        ("D", "Signature", 20),
        ("E", "Date", 14),
        ("F", "Comments", 35),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    approvals = [
        ("NDA Register Complete", "HR Manager / Contracts Manager"),
        ("Expiration Review", "Information Security Manager"),
        ("Overall Approval", "CISO"),
    ]

    for row_idx, (approval_type, role) in enumerate(approvals, start=3):
        ws[f"A{row_idx}"] = approval_type
        ws[f"B{row_idx}"] = role
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            if col in ["C", "D", "E", "F"]:
                ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def main() -> int:
    logger.info("=" * 60)
    logger.info("Generating %s", OUTPUT_FILENAME)
    logger.info("=" * 60)

    try:
        wb = create_workbook()
        styles = setup_styles()

        create_instructions_sheet(wb["Instructions"], styles)
        create_active_ndas_sheet(wb["Active_NDAs"], styles)
        create_signatory_register_sheet(wb["Signatory_Register"], styles)
        create_expiration_monitor_sheet(wb["Expiration_Monitor"], styles)
        create_renewal_tracking_sheet(wb["Renewal_Tracking"], styles)
        create_evidence_register_sheet(wb["Evidence_Register"], styles)
        create_approval_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)
        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        return 0

    except Exception as e:
        logger.error("FAILED: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# =============================================================================
