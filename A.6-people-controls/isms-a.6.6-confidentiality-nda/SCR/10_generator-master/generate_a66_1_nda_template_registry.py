#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.6.1 - NDA Template Registry & Inventory Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.6: Confidentiality or Non-Disclosure Agreements

Assessment Domain 1 of 4: NDA Template Registry & Inventory

This script generates a workbook for managing NDA templates, versions, and
applicability across different stakeholder categories.

**Generated Workbook Structure:**
1. Instructions - Template registry guidance
2. Template_Registry - Master list of NDA templates
3. Template_Versions - Version history tracking
4. Applicability_Matrix - When to use which template
5. Clause_Library - Standard confidentiality clauses
6. Evidence_Register - Audit evidence tracking
7. Approval_SignOff - Stakeholder approval

================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

DOCUMENT_ID = "ISMS-IMP-A.6.6.1"
WORKBOOK_NAME = "NDA Template Registry and Inventory"
CONTROL_ID = "A.6.6"
CONTROL_NAME = "Confidentiality or Non-Disclosure Agreements"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


def setup_styles():
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
    }


def create_workbook() -> Workbook:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    for name in ["Instructions", "Template_Registry", "Template_Versions",
                 "Applicability_Matrix", "Clause_Library",
                 "Evidence_Register", "Approval_SignOff"]:
        wb.create_sheet(title=name)
    return wb


def create_instructions_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:H3")
    ws["A3"] = CONTROL_REF
    ws["A3"].font = Font(bold=True, size=12)

    instructions = [
        "",
        "PURPOSE",
        "This workbook maintains the registry of NDA templates, version history, and applicability",
        "guidance for confidentiality agreements across all stakeholder categories.",
        "",
        "SCOPE",
        "- All NDA and confidentiality agreement templates",
        "- Version control and change history",
        "- Applicability matrix by stakeholder type",
        "- Standard clause library for consistency",
        "",
        "COMPLETION STEPS",
        "1. Template_Registry: Register all NDA templates with metadata",
        "2. Template_Versions: Track version history and changes",
        "3. Applicability_Matrix: Define when each template applies",
        "4. Clause_Library: Document standard confidentiality clauses",
        "5. Evidence_Register: Link supporting evidence",
        "6. Approval_SignOff: Obtain required approvals",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, text in enumerate(instructions, start=5):
        ws[f"A{i}"] = text
        if text in ["PURPOSE", "SCOPE", "COMPLETION STEPS"]:
            ws[f"A{i}"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 80


def create_template_registry_sheet(ws, styles):
    ws.merge_cells("A1:L1")
    ws["A1"] = "NDA Template Registry"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Template_ID", 14),
        ("B", "Template_Name", 35),
        ("C", "Template_Type", 18),
        ("D", "Stakeholder_Category", 22),
        ("E", "Current_Version", 14),
        ("F", "Effective_Date", 14),
        ("G", "Legal_Review_Date", 16),
        ("H", "Legal_Reviewer", 20),
        ("I", "Owner", 20),
        ("J", "Storage_Location", 30),
        ("K", "Status", 14),
        ("L", "Notes", 30),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Data validations
    type_val = DataValidation(type="list", formula1='"Standard NDA,Mutual NDA,One-Way NDA,Employment,Contractor,Vendor,Customer,Partner"')
    type_val.add("C3:C100")
    ws.add_data_validation(type_val)

    category_val = DataValidation(type="list", formula1='"Employees,Contractors,Consultants,Vendors,Suppliers,Partners,Customers,Board Members,Visitors"')
    category_val.add("D3:D100")
    ws.add_data_validation(category_val)

    status_val = DataValidation(type="list", formula1='"Active,Draft,Under Review,Superseded,Archived"')
    status_val.add("K3:K100")
    ws.add_data_validation(status_val)

    for row in range(3, 23):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_template_versions_sheet(ws, styles):
    ws.merge_cells("A1:J1")
    ws["A1"] = "Template Version History"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Template_ID", 14),
        ("B", "Version", 12),
        ("C", "Version_Date", 14),
        ("D", "Change_Description", 45),
        ("E", "Change_Reason", 30),
        ("F", "Changed_By", 20),
        ("G", "Legal_Approved", 14),
        ("H", "Legal_Approver", 20),
        ("I", "Approval_Date", 14),
        ("J", "Notes", 30),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    yes_no = DataValidation(type="list", formula1='"Yes,No,Pending"')
    yes_no.add("G3:G100")
    ws.add_data_validation(yes_no)

    for row in range(3, 23):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_applicability_matrix_sheet(ws, styles):
    ws.merge_cells("A1:I1")
    ws["A1"] = "NDA Applicability Matrix"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Stakeholder_Category", 22),
        ("B", "Access_Type", 25),
        ("C", "Information_Classification", 22),
        ("D", "Required_Template", 25),
        ("E", "Timing", 20),
        ("F", "Duration", 18),
        ("G", "Post_Termination", 18),
        ("H", "Mandatory", 12),
        ("I", "Notes", 35),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Pre-populate common scenarios
    scenarios = [
        ("Employees", "Full system access", "All levels", "Employment NDA", "Before start date", "Indefinite", "2 years", "Yes"),
        ("Contractors", "Project-specific access", "Confidential+", "Contractor NDA", "Before access granted", "Contract duration", "2 years", "Yes"),
        ("Consultants", "Advisory access", "Confidential+", "Consultant NDA", "Before engagement", "Engagement duration", "2 years", "Yes"),
        ("Vendors", "System integration", "Confidential", "Vendor NDA", "Before contract signing", "Contract duration", "3 years", "Yes"),
        ("Suppliers", "Limited access", "Internal", "Supplier NDA", "Before access granted", "Contract duration", "1 year", "Yes"),
        ("Partners", "Mutual sharing", "Confidential+", "Mutual NDA", "Before discussions", "Partnership duration", "3 years", "Yes"),
        ("Customers", "Product access", "Confidential", "Customer NDA", "Before disclosure", "Agreement duration", "2 years", "Conditional"),
        ("Board Members", "Strategic access", "All levels", "Director NDA", "Before appointment", "Tenure + 3 years", "3 years", "Yes"),
        ("Visitors", "Escorted access", "Internal only", "Visitor NDA", "Before site entry", "Visit duration", "1 year", "Conditional"),
    ]

    for row_idx, scenario in enumerate(scenarios, start=3):
        for col_idx, value in enumerate(scenario):
            col = get_column_letter(col_idx + 1)
            ws[f"{col}{row_idx}"] = value
            ws[f"{col}{row_idx}"].border = styles["border"]

    mandatory_val = DataValidation(type="list", formula1='"Yes,No,Conditional"')
    mandatory_val.add("H3:H50")
    ws.add_data_validation(mandatory_val)

    ws.freeze_panes = "A3"


def create_clause_library_sheet(ws, styles):
    ws.merge_cells("A1:G1")
    ws["A1"] = "Standard Confidentiality Clause Library"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Clause_ID", 12),
        ("B", "Clause_Name", 30),
        ("C", "Clause_Category", 20),
        ("D", "Clause_Purpose", 40),
        ("E", "Standard_Text", 60),
        ("F", "Mandatory", 12),
        ("G", "Notes", 30),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Pre-populate common clauses
    clauses = [
        ("CL-001", "Definition of Confidential Information", "Definitions", "Define what constitutes confidential information", "[Organisation-specific definition text]", "Yes"),
        ("CL-002", "Obligations of Receiving Party", "Obligations", "Specify recipient's duties", "[Standard obligation text]", "Yes"),
        ("CL-003", "Permitted Disclosures", "Exceptions", "Define allowed disclosures", "[Standard exceptions text]", "Yes"),
        ("CL-004", "Return of Information", "Termination", "Require return/destruction of info", "[Standard return clause]", "Yes"),
        ("CL-005", "Duration of Obligations", "Term", "Specify confidentiality period", "[Duration specification]", "Yes"),
        ("CL-006", "Injunctive Relief", "Remedies", "Allow equitable remedies", "[Standard injunction text]", "Recommended"),
        ("CL-007", "No License Granted", "IP Rights", "Clarify no IP transfer", "[Standard no-license clause]", "Recommended"),
        ("CL-008", "Governing Law", "Legal", "Specify applicable law", "[Jurisdiction specification]", "Yes"),
    ]

    for row_idx, clause in enumerate(clauses, start=3):
        for col_idx, value in enumerate(clause):
            col = get_column_letter(col_idx + 1)
            ws[f"{col}{row_idx}"] = value
            ws[f"{col}{row_idx}"].border = styles["border"]

    category_val = DataValidation(type="list", formula1='"Definitions,Obligations,Exceptions,Term,Termination,Remedies,IP Rights,Legal,General"')
    category_val.add("C3:C50")
    ws.add_data_validation(category_val)

    mandatory_val = DataValidation(type="list", formula1='"Yes,No,Recommended"')
    mandatory_val.add("F3:F50")
    ws.add_data_validation(mandatory_val)

    ws.freeze_panes = "A3"


def create_evidence_register_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "Evidence Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Evidence_ID", 14),
        ("B", "Template_Ref", 14),
        ("C", "Evidence_Type", 20),
        ("D", "Description", 40),
        ("E", "Storage_Location", 35),
        ("F", "Collected_Date", 14),
        ("G", "Collected_By", 20),
        ("H", "Retention_Until", 14),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    type_val = DataValidation(type="list", formula1='"Template Document,Legal Review,Version History,Approval Record,Training Material,Other"')
    type_val.add("C3:C50")
    ws.add_data_validation(type_val)

    for row in range(3, 23):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_approval_sheet(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "Approval and Sign-Off"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Approval_Type", 25),
        ("B", "Approver_Role", 25),
        ("C", "Approver_Name", 25),
        ("D", "Signature", 20),
        ("E", "Date", 14),
        ("F", "Comments", 35),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    approvals = [
        ("Template Registry Complete", "Information Security Manager"),
        ("Legal Review Complete", "Legal Counsel"),
        ("Overall Approval", "CISO"),
    ]

    for row_idx, (approval_type, role) in enumerate(approvals, start=3):
        ws[f"A{row_idx}"] = approval_type
        ws[f"B{row_idx}"] = role
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            if col in ["C", "D", "E", "F"]:
                ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def main() -> int:
    logger.info("=" * 60)
    logger.info("Generating %s", OUTPUT_FILENAME)
    logger.info("=" * 60)

    try:
        wb = create_workbook()
        styles = setup_styles()

        create_instructions_sheet(wb["Instructions"], styles)
        create_template_registry_sheet(wb["Template_Registry"], styles)
        create_template_versions_sheet(wb["Template_Versions"], styles)
        create_applicability_matrix_sheet(wb["Applicability_Matrix"], styles)
        create_clause_library_sheet(wb["Clause_Library"], styles)
        create_evidence_register_sheet(wb["Evidence_Register"], styles)
        create_approval_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)
        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        return 0

    except Exception as e:
        logger.error("FAILED: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# =============================================================================
