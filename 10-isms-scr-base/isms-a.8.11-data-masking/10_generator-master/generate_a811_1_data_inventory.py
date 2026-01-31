#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# Licensed under AGPL-3.0-or-later with commercial licensing option
#
# This file is part of the ISMS Compliance Framework
# See /LICENSE for full terms and /LICENSES/COMMERCIAL.md for commercial options
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.11.1 - Data Inventory & Classification Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.11: Data Masking
Assessment Domain 1 of 5: Data Inventory & Classification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific data systems, classification taxonomies, and
masking requirements.

Key customization areas:
1. System inventory (match your actual databases and applications)
2. Data category taxonomy (adapt to your classification scheme)
3. Regulatory mappings (GDPR, FADP, HIPAA per your jurisdictions)
4. Masking priority criteria (based on your risk appetite)
5. Evidence requirements (aligned with your audit framework)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for data masking)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for documenting
and classifying sensitive data across organizational systems.

**Purpose:**
Enables systematic inventory of sensitive data fields requiring masking,
supporting evidence-based classification and prioritization.

**Assessment Scope:**
- System and database inventory
- Sensitive data field identification
- Data category classification (PII, financial, health, etc.)
- Regulatory requirement mapping (GDPR, FADP, HIPAA, etc.)
- Data ownership assignment
- Masking priority determination
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and data taxonomy
2. System Inventory - Master list of systems/databases
3. Data Category Reference - Standard data categories (read-only)
4. Sensitive Data Inventory - Field-level data catalog
5. Classification Matrix - Data sensitivity classification
6. Regulatory Mapping - Compliance requirement mapping
7. Data Owner Assignment - Ownership and accountability
8. Masking Priority Matrix - Risk-based prioritization
9. Gap Analysis - Incomplete classifications and missing data
10. Evidence Register - Audit evidence tracking
11. Summary Dashboard - Consolidated metrics and status

**Integration:**
This assessment feeds into ISMS-IMP-A.8.11.5 Compliance Dashboard, which
consolidates data from all five assessment domains for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a811_1_data_inventory.py

Output:
    File: ISMS_IMP_A_8_11_1_Data_Inventory_Assessment_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Inventory all systems containing data
    2. Identify sensitive data fields at table/column level
    3. Apply data category classifications
    4. Map to regulatory requirements
    5. Assign data owners
    6. Determine masking priorities
    7. Collect audit evidence
    8. Obtain stakeholder approvals
    9. Feed results into A.8.11.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.11
Assessment Domain:    1 of 5 (Data Inventory & Classification)
Related Policy:       ISMS-POL-A.8.11 (Data Masking Policy)
Script Version:       1.0
Python Version:       3.8+

Related Documents:
    - ISMS-POL-A.8.11: Data Masking Policy
    - ISMS-IMP-A.8.11.1: Data Inventory & Classification Assessment Guide
    - ISMS-IMP-A.8.11.2: Masking Technique Selection (Domain 2)
    - ISMS-IMP-A.8.11.3: Environment Coverage (Domain 3)
    - ISMS-IMP-A.8.11.4: Testing & Validation (Domain 4)
    - ISMS-IMP-A.8.11.5: Compliance Dashboard (Consolidation)

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)



# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.11.1"
WORKBOOK_NAME = "Data Inventory & Classification Assessment"
CONTROL_ID = "A.8.11"
CONTROL_NAME = "Data Masking"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •



# ============================================================================
# SECTION 1: CONSTANTS & CONFIGURATION
# ============================================================================

# Document Information
WORKBOOK_VERSION = "1.0"
CONTROL_ID = "A.8.11"
WORKBOOK_ID = "ISMS-IMP-A.8.11.1"
RELATED_POLICY = "ISMS-POL-A.8.11-S2.1"
ASSESSMENT_AREA = "Data Inventory & Classification"

# Color Scheme (consistent with A.8.23/8.24)
COLOR_HEADER = "003366"          # Dark blue
COLOR_SUBHEADER = "4472C4"       # Medium blue
COLOR_COLUMN_HEADER = "D9D9D9"   # Light gray
COLOR_INPUT = "FFFFCC"           # Light yellow (user input)
COLOR_INFO = "E7E6E6"            # Light gray (info/example rows)
COLOR_COMPLETE = "C6EFCE"        # Light green (\u2705)
COLOR_PARTIAL = "FFEB9C"         # Light yellow (\u26A0\uFE0F)
COLOR_MISSING = "FFC7CE"         # Light red (\u274C)
COLOR_PLANNED = "B4C7E7"         # Light blue (\u1F4CB)

# Standard column widths
WIDTH_NARROW = 12
WIDTH_MEDIUM = 20
WIDTH_WIDE = 25
WIDTH_EXTRA_WIDE = 30
WIDTH_DESCRIPTION = 35


# ============================================================================
# SECTION 2: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create all sheets in order
    sheets = [
        "Instructions_Legend",
        "System_Inventory",
        "Data_Category_Reference",
        "Sensitive_Data_Inventory",
        "Classification_Matrix",
        "Regulatory_Mapping",
        "Data_Owner_Assignment",
        "Masking_Priority_Matrix",
        "Gap_Analysis",
        "Evidence_Register",
        "Summary_Dashboard",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    
    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style TEMPLATES (dictionaries), not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color=COLOR_COLUMN_HEADER, end_color=COLOR_COLUMN_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "info_cell": {
            "fill": PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"),
            "font": Font(name="Calibri", size=10, italic=True),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_complete": {
            "fill": PatternFill(start_color=COLOR_COMPLETE, end_color=COLOR_COMPLETE, fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color=COLOR_PARTIAL, end_color=COLOR_PARTIAL, fill_type="solid")
        },
        "status_missing": {
            "fill": PatternFill(start_color=COLOR_MISSING, end_color=COLOR_MISSING, fill_type="solid")
        },
        "status_planned": {
            "fill": PatternFill(start_color=COLOR_PLANNED, end_color=COLOR_PLANNED, fill_type="solid")
        },
    }
    
    return styles


def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            italic=getattr(style_dict["font"], 'italic', False),
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    These are added to worksheet once, then applied to multiple cells.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_unknown': DataValidation(
            type="list",
            formula1='"Yes,No,Unknown"',
            allow_blank=False
        ),
        'yes_no_partial': DataValidation(
            type="list",
            formula1='"Yes,No,Partial"',
            allow_blank=False
        ),
        'yes_no_partial_planned_na': DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Planned,N/A"',
            allow_blank=False
        ),
        'yes_no_conditional': DataValidation(
            type="list",
            formula1='"Yes,No,Conditional,Unknown"',
            allow_blank=False
        ),
        'status_icons': DataValidation(
            type="list",
            formula1='"\u2705 Complete,\u26A0\uFE0F Partial,\u274C Missing,\u1F4CB Planned,N/A"',
            allow_blank=False
        ),
        'system_type': DataValidation(
            type="list",
            formula1='"Database,Application,SaaS,File Share,API,Data Warehouse,Backup System,Archive,Other"',
            allow_blank=False
        ),
        'environment_type': DataValidation(
            type="list",
            formula1='"Production,Development,Test/QA,UAT,Training,Analytics,DR/Backup,Decommissioned"',
            allow_blank=False
        ),
        'hosting_location': DataValidation(
            type="list",
            formula1='"On-Premises,AWS,Azure,GCP,Multi-Cloud,Hybrid,Third-Party Hosted,Unknown"',
            allow_blank=False
        ),
        'data_type': DataValidation(
            type="list",
            formula1='"String,Integer,Date,Boolean,Binary,JSON,Other"',
            allow_blank=False
        ),
        'data_category': DataValidation(
            type="list",
            formula1='"CAT-PII-D,CAT-PII-I,CAT-FIN,CAT-HLT,CAT-CRD,CAT-PRP,CAT-LOC,CAT-BIO,CAT-GEN,CAT-CHD"',
            allow_blank=False
        ),
        'pii_type': DataValidation(
            type="list",
            formula1='"Name,SSN,Email,Phone,Address,DOB,PAN,Medical Record,Password,IP Address,Other"',
            allow_blank=False
        ),
        'sensitivity_level': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,Public"',
            allow_blank=False
        ),
        'masking_status': DataValidation(
            type="list",
            formula1='"Masked,Not Masked,Partially Masked,Encrypted,Planned"',
            allow_blank=False
        ),
        'discovery_method': DataValidation(
            type="list",
            formula1='"Automated Scan,Schema Review,Manual Sample,Data Owner Interview,Unknown"',
            allow_blank=False
        ),
        'org_classification': DataValidation(
            type="list",
            formula1='"Restricted,Confidential,Internal,Public"',
            allow_blank=False
        ),
        'gdpr_article': DataValidation(
            type="list",
            formula1='"Art. 4 (PII),Art. 9 (Special),Art. 32 (Security),N/A"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'exposure_risk': DataValidation(
            type="list",
            formula1='"Very High,High,Medium,Low,Very Low"',
            allow_blank=False
        ),
        'priority_tier': DataValidation(
            type="list",
            formula1='"P1,P2,P3,P4"',
            allow_blank=False
        ),
        'implementation_status': DataValidation(
            type="list",
            formula1='"Not Started,In Progress,Complete,Blocked"',
            allow_blank=False
        ),
        'gap_category': DataValidation(
            type="list",
            formula1='"Inventory Missing,Classification Missing,Ownership Missing,Masking Missing,Regulatory Unknown,Other"',
            allow_blank=False
        ),
        'gap_status': DataValidation(
            type="list",
            formula1='"Open,In Progress,Complete,Accepted Risk"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Data Discovery Report,Schema Documentation,DPIA,Data Flow Diagram,Classification Review,RACI Matrix,Meeting Minutes,Approval Email,Tool Report,Other"',
            allow_blank=False
        ),
        'review_frequency': DataValidation(
            type="list",
            formula1='"Annual,Semi-Annual,Quarterly,As-Needed,N/A"',
            allow_blank=False
        ),
        'confidentiality_level': DataValidation(
            type="list",
            formula1='"Public,Internal,Confidential,Restricted"',
            allow_blank=False
        ),
    }
    
    # Add all validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    return validations


# ============================================================================
# SECTION 4: SHEET 1 - INSTRUCTIONS & LEGEND
# ============================================================================

def create_instructions_legend(ws, styles):
    """Create Instructions & Legend sheet."""
    
    # HEADER
    ws.merge_cells('A1:H1')
    header = ws['A1']
    header.value = f"{WORKBOOK_ID} — {ASSESSMENT_AREA}"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 30
    
    # SUBTITLE
    ws.merge_cells('A2:H2')
    subtitle = ws['A2']
    subtitle.value = "ISO/IEC 27001:2022 - Control A.8.11: Data Masking"
    apply_style(subtitle, styles["subheader"])
    ws.row_dimensions[2].height = 20
    
    # DOCUMENT INFORMATION BLOCK
    row = 4
    info_fields = [
        ("Document ID:", WORKBOOK_ID),
        ("Assessment Area:", ASSESSMENT_AREA),
        ("Related Policy:", RELATED_POLICY),
        ("Version:", WORKBOOK_VERSION),
        ("Assessment Date:", "[USER INPUT]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Quarterly (Inventory) / Annual (Classification Review)"),
    ]
    
    for label, value in info_fields:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if "[USER INPUT]" in value:
            apply_style(ws[f'B{row}'], styles["input_cell"])
        row += 1
    
    # HOW TO USE THIS WORKBOOK
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "HOW TO USE THIS WORKBOOK"
    ws[f'A{row}'].font = Font(size=11, bold=True)
    row += 1
    
    instructions = [
        "1. Start with System_Inventory — list ALL systems/databases that process data",
        "2. Use Data_Category_Reference to understand sensitivity taxonomy",
        "3. Complete Sensitive_Data_Inventory for each system (table/field level)",
        "4. Apply classification using the Classification_Matrix sheet",
        "5. Map data to regulatory requirements in Regulatory_Mapping",
        "6. Assign data owners in Data_Owner_Assignment",
        "7. Prioritize masking efforts in Masking_Priority_Matrix",
        "8. Document gaps in Gap_Analysis",
        "9. Maintain evidence in Evidence_Register",
        "10. Review summary metrics in Summary_Dashboard",
        "11. Obtain required approvals",
    ]
    
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 15
        row += 1
    
    # DATA CATEGORY TAXONOMY
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "DATA CATEGORY TAXONOMY (Quick Reference)"
    ws[f'A{row}'].font = Font(size=11, bold=True)
    row += 1
    
    # Category table headers
    ws[f'A{row}'] = "Category ID"
    ws[f'B{row}'] = "Category Name"
    ws[f'C{row}'] = "Examples"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f'{col}{row}'], styles["column_header"])
    row += 1
    
    categories = [
        ("CAT-PII-D", "Direct PII", "Name, SSN/AHV, Passport, Email, Phone"),
        ("CAT-PII-I", "Indirect PII", "DOB+ZIP, Job Title+Department"),
        ("CAT-FIN", "Financial Data", "Credit Card (PAN), IBAN, Account Balance"),
        ("CAT-HLT", "Health Data", "Diagnoses, Prescriptions, Medical Records"),
        ("CAT-CRD", "Credentials", "Passwords, API Keys, Tokens"),
        ("CAT-PRP", "Proprietary", "Trade Secrets, Pricing, IP"),
        ("CAT-LOC", "Location Data", "GPS, IP Address, Travel History"),
        ("CAT-BIO", "Biometric", "Fingerprints, Facial Geometry"),
        ("CAT-GEN", "Genetic Data", "DNA, Genetic Test Results"),
        ("CAT-CHD", "Child Data", "Data of minors (<16 GDPR / <18)"),
    ]
    
    for cat_id, cat_name, examples in categories:
        ws[f'A{row}'] = cat_id
        ws[f'B{row}'] = cat_name
        ws[f'C{row}'] = examples
        ws.row_dimensions[row].height = 15
        row += 1
    
    # SENSITIVITY CLASSIFICATION LEVELS
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "SENSITIVITY CLASSIFICATION LEVELS"
    ws[f'A{row}'].font = Font(size=11, bold=True)
    row += 1
    
    ws[f'A{row}'] = "Level"
    ws[f'B{row}'] = "Definition"
    ws[f'C{row}'] = "Masking Requirement"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f'{col}{row}'], styles["column_header"])
    row += 1
    
    levels = [
        ("Critical", "Severe harm if exposed, regulatory breach", "SHALL mask in ALL non-production"),
        ("High", "Substantial harm, privacy violation", "SHALL mask in non-production"),
        ("Medium", "Moderate harm, business impact", "SHOULD mask in non-production"),
        ("Low", "Minimal harm", "MAY mask based on risk"),
        ("Public", "No confidentiality requirement", "N/A"),
    ]
    
    for level, definition, requirement in levels:
        ws[f'A{row}'] = level
        ws[f'B{row}'] = definition
        ws[f'C{row}'] = requirement
        ws.row_dimensions[row].height = 15
        row += 1
    
    # STATUS LEGEND
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "STATUS LEGEND"
    ws[f'A{row}'].font = Font(size=11, bold=True)
    row += 1
    
    ws[f'A{row}'] = "Symbol"
    ws[f'B{row}'] = "Status"
    ws[f'C{row}'] = "Description"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f'{col}{row}'], styles["column_header"])
    row += 1
    
    statuses = [
        ("\u2705", "Complete", "Inventory complete, classified"),
        ("\u26A0\uFE0F", "Partial", "Partial inventory/classification"),
        ("\u274C", "Missing", "Not inventoried/classified"),
        ("\u1F4CB", "Planned", "Discovery/classification planned"),
        ("N/A", "Not Applicable", "System contains no sensitive data"),
    ]
    
    for symbol, status, description in statuses:
        ws[f'A{row}'] = symbol
        ws[f'B{row}'] = status
        ws[f'C{row}'] = description
        row += 1
    
    # ACCEPTABLE EVIDENCE
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "ACCEPTABLE EVIDENCE (Examples)"
    ws[f'A{row}'].font = Font(size=11, bold=True)
    row += 1
    
    evidence_examples = [
        "✓ Database schema documentation",
        "✓ Data dictionaries with field definitions",
        "✓ Data discovery tool reports (e.g., BigID, OneTrust, etc.)",
        "✓ Data flow diagrams",
        "✓ Data Processing Impact Assessments (DPIA)",
        "✓ System design documents",
        "✓ Privacy policy documentation",
        "✓ Data retention schedules",
        "✓ Data owner assignment matrices (RACI)",
        "✓ Classification review meeting minutes",
        "✓ Regulatory compliance assessments",
        "✓ Third-party data sharing agreements",
    ]
    
    for evidence in evidence_examples:
        ws[f'A{row}'] = evidence
        ws.row_dimensions[row].height = 15
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_WIDE
    ws.column_dimensions['C'].width = WIDTH_EXTRA_WIDE
    
    # Freeze panes
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 5: SHEET 2 - SYSTEM INVENTORY
# ============================================================================

def create_system_inventory(ws, styles):
    """Create System Inventory sheet with 50-row template."""
    
    validations = create_base_validations(ws)
    
    # HEADER
    ws.merge_cells('A1:Q1')
    header = ws['A1']
    header.value = "SYSTEM & DATABASE INVENTORY"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 25
    
    # SUBTITLE
    ws.merge_cells('A2:Q2')
    subtitle = ws['A2']
    subtitle.value = "List ALL systems that process, store, or transmit data (50 row template)"
    apply_style(subtitle, styles["subheader"])
    ws.row_dimensions[2].height = 20
    
    # ASSESSMENT QUESTION
    ws.merge_cells('A3:Q3')
    ws['A3'] = "Does your organization maintain a comprehensive inventory of all systems containing sensitive data?"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 25
    
    # RESPONSE DROPDOWN
    ws['A4'] = "Response:"
    ws['A4'].font = Font(bold=True)
    ws['B4'].value = ""
    apply_style(ws['B4'], styles["input_cell"])
    validations['yes_no_partial_planned_na'].add(ws['B4'])
    
    # COLUMN HEADERS (Row 6)
    headers = [
        ("A", "System ID", 15),
        ("B", "System Name", 25),
        ("C", "System Type", 18),
        ("D", "Environment", 15),
        ("E", "Contains Sensitive Data?", 18),
        ("F", "Data Categories Present", 25),
        ("G", "Hosting Location", 20),
        ("H", "Primary Data Owner", 20),
        ("I", "System Owner/Admin", 20),
        ("J", "Inventory Status", 18),
        ("K", "Last Inventory Date", 15),
        ("L", "Next Review Date", 15),
        ("M", "Record Count (Approx)", 15),
        ("N", "Retention Period", 15),
        ("O", "Regulatory Scope", 20),
        ("P", "Decommission Date", 15),
        ("Q", "Notes/Comments", 30),
    ]
    
    for col, header_text, width in headers:
        cell = ws[f'{col}6']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # EXAMPLE ROW (Row 7) - Gray italic
    example_data = [
        "SYS-001",
        "Customer CRM System",
        "Application",
        "Production",
        "Yes",
        "CAT-PII-D, CAT-FIN",
        "AWS",
        "VP Sales",
        "IT Manager",
        "\u2705 Complete",
        "2024-12-15",
        "2025-03-15",
        "500000",
        "7 years",
        "GDPR, FADP",
        "",
        "Primary customer database",
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=7, column=col_idx)
        cell.value = value
        apply_style(cell, styles["info_cell"])
    
    # DATA ENTRY ROWS (8-57: 50 rows)
    for row_idx in range(8, 58):
        for col_idx in range(1, 18):  # A-Q = 17 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Apply data validations to appropriate columns
    for row in range(8, 58):
        validations['system_type'].add(ws[f'C{row}'])
        validations['environment_type'].add(ws[f'D{row}'])
        validations['yes_no_unknown'].add(ws[f'E{row}'])
        validations['hosting_location'].add(ws[f'G{row}'])
        validations['status_icons'].add(ws[f'J{row}'])
    
    # CHECKLIST SECTION (Starting Row 59)
    checklist_row = 59
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "SYSTEM INVENTORY CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    # Checklist headers
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Is a system inventory maintained?",
        "Are all production systems documented?",
        "Are non-production environments included?",
        "Are SaaS/cloud systems included?",
        "Are backup/archive systems included?",
        "Are data warehouses/analytics systems included?",
        "Is system type classified for each entry?",
        "Is hosting location documented?",
        "Are system owners assigned?",
        "Are data owners assigned?",
        "Is inventory reviewed quarterly?",
        "Are decommissioned systems tracked?",
        "Is record count estimated per system?",
        "Is retention period documented?",
        "Are regulatory requirements mapped?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        ws.row_dimensions[checklist_row].height = 20
        checklist_row += 1
    
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    
    # REFERENCE TABLE: System Types (Starting Row 76)
    ref_row = 76
    ws.merge_cells(f'A{ref_row}:C{ref_row}')
    ws[f'A{ref_row}'] = "SYSTEM TYPE DEFINITIONS"
    ws[f'A{ref_row}'].font = Font(bold=True, size=11)
    ref_row += 1
    
    ws[f'A{ref_row}'] = "System Type"
    ws[f'B{ref_row}'] = "Description"
    ws[f'C{ref_row}'] = "Examples"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f'{col}{ref_row}'], styles["column_header"])
    ref_row += 1
    
    system_types = [
        ("Database", "Relational or NoSQL databases", "Oracle, PostgreSQL, MongoDB, SQL Server"),
        ("Application", "Business applications", "CRM, ERP, HR systems, custom apps"),
        ("SaaS", "Cloud-based software services", "Salesforce, Workday, ServiceNow"),
        ("File Share", "File storage systems", "SharePoint, NAS, file servers"),
        ("API", "API gateways and services", "REST APIs, integration platforms"),
        ("Data Warehouse", "Analytics and BI systems", "Snowflake, Redshift, BigQuery"),
        ("Backup System", "Backup and recovery systems", "Veeam, backup appliances"),
        ("Archive", "Long-term storage", "Tape systems, cold storage"),
        ("Other", "Other system types", "(Document in notes)"),
    ]
    
    for sys_type, description, examples in system_types:
        ws[f'A{ref_row}'] = sys_type
        ws[f'B{ref_row}'] = description
        ws[f'C{ref_row}'] = examples
        ws.row_dimensions[ref_row].height = 15
        ref_row += 1
    
    # Freeze panes
    ws.freeze_panes = "A7"


# ============================================================================
# SECTION 6: SHEET 3 - DATA CATEGORY REFERENCE
# ============================================================================

def create_data_category_reference(ws, styles):
    """Create read-only Data Category Reference sheet."""
    
    # HEADER
    ws.merge_cells('A1:G1')
    header = ws['A1']
    header.value = "SENSITIVE DATA CATEGORY REFERENCE"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 25
    
    # SUBTITLE
    ws.merge_cells('A2:G2')
    subtitle = ws['A2']
    subtitle.value = "Reference taxonomy from ISMS-POL-A.8.11-S2.1"
    apply_style(subtitle, styles["subheader"])
    ws.row_dimensions[2].height = 20
    
    # COLUMN HEADERS (Row 4)
    headers = [
        ("A", "Category ID", 15),
        ("B", "Category Name", 25),
        ("C", "Description", 40),
        ("D", "Examples", 50),
        ("E", "Sensitivity Level", 15),
        ("F", "Masking Priority", 18),
        ("G", "Regulatory Driver", 25),
    ]
    
    for col, header_text, width in headers:
        cell = ws[f'{col}4']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # DATA CATEGORY DEFINITIONS (Pre-populated from Policy S2.1)
    row = 5
    categories_data = [
        ("CAT-PII-D", "Direct PII", "Data that directly identifies an individual",
         "Full Name, SSN/AHV, Passport, Driver's License, Email, Phone, Address, Photo",
         "High/Critical", "Required", "FADP, GDPR"),
        
        ("CAT-PII-I", "Indirect PII", "Data that can identify when combined",
         "DOB+ZIP, Job Title+Department, Employee ID",
         "Medium/High", "Required when combined", "FADP, GDPR"),
        
        ("CAT-FIN", "Financial Data", "Payment and financial account data",
         "PAN (Credit Card), CVV, IBAN, Account Balance, Salary, Tax ID",
         "Critical", "Required", "PCI-DSS, FADP, GDPR"),
        
        ("CAT-HLT", "Health Data", "Medical and health information",
         "Medical Record Number, Diagnoses, Medications, Lab Results, Insurance ID, Mental Health Records",
         "Critical", "Required", "FADP, GDPR Article 9"),
        
        ("CAT-CRD", "Credentials", "Authentication and access data",
         "Passwords, Password Hashes, API Keys, OAuth Tokens, Private Keys, DB Connection Strings",
         "Critical", "Required (Never display)", "Best Practice"),
        
        ("CAT-PRP", "Proprietary", "Business-sensitive information",
         "Trade Secrets, Pricing Models, Strategic Plans, Source Code",
         "High", "Recommended", "Business Confidentiality"),
        
        ("CAT-LOC", "Location Data", "Geographic and tracking data",
         "GPS Coordinates, IP Addresses, Travel History",
         "Medium/High", "Recommended", "GDPR (personal location)"),
        
        ("CAT-BIO", "Biometric Data", "Physical/behavioral identifiers",
         "Fingerprints, Facial Geometry, Voice Prints, Retina Scans",
         "Critical", "Required", "GDPR Article 9"),
        
        ("CAT-GEN", "Genetic Data", "DNA and genetic information",
         "DNA Sequences, Genetic Test Results, Hereditary Information",
         "Critical", "Required", "GDPR Article 9"),
        
        ("CAT-CHD", "Child Data", "Data relating to minors",
         "Any PII of persons under 16 (GDPR) / 18 (local)",
         "Critical", "Required", "GDPR Article 8, COPPA"),
    ]
    
    for cat_id, cat_name, description, examples, sensitivity, masking, regulations in categories_data:
        ws[f'A{row}'] = cat_id
        ws[f'B{row}'] = cat_name
        ws[f'C{row}'] = description
        ws[f'D{row}'] = examples
        ws[f'E{row}'] = sensitivity
        ws[f'F{row}'] = masking
        ws[f'G{row}'] = regulations
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{row}'].alignment = Alignment(wrap_text=True)
        
        ws.row_dimensions[row].height = 30
        row += 1
    
    # Freeze panes
    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 7: SHEET 4 - SENSITIVE DATA INVENTORY
# ============================================================================

def create_sensitive_data_inventory(ws, styles):
    """Create Sensitive Data Inventory sheet with 100-row template."""
    
    validations = create_base_validations(ws)
    
    # HEADER
    ws.merge_cells('A1:R1')
    header = ws['A1']
    header.value = "SENSITIVE DATA ELEMENT INVENTORY"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 25
    
    # SUBTITLE
    ws.merge_cells('A2:R2')
    subtitle = ws['A2']
    subtitle.value = "Document sensitive data at the table/field level (100 row template)"
    apply_style(subtitle, styles["subheader"])
    ws.row_dimensions[2].height = 20
    
    # ASSESSMENT QUESTION
    ws.merge_cells('A3:R3')
    ws['A3'] = "Has your organization identified all sensitive data fields across systems?"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 25
    
    # RESPONSE DROPDOWN
    ws['A4'] = "Response:"
    ws['A4'].font = Font(bold=True)
    ws['B4'].value = ""
    apply_style(ws['B4'], styles["input_cell"])
    validations['yes_no_partial_planned_na'].add(ws['B4'])
    
    # COLUMN HEADERS (Row 6)
    headers = [
        ("A", "Data Element ID", 15),
        ("B", "System Name", 25),
        ("C", "Database/Schema", 20),
        ("D", "Table/Collection Name", 25),
        ("E", "Field/Column Name", 25),
        ("F", "Data Type", 15),
        ("G", "Data Category", 15),
        ("H", "Specific PII Type", 20),
        ("I", "Sensitivity Level", 15),
        ("J", "Contains PII?", 12),
        ("K", "Regulatory Scope", 20),
        ("L", "Masking Required?", 15),
        ("M", "Current Masking Status", 18),
        ("N", "Record Count (Approx)", 15),
        ("O", "Data Owner", 20),
        ("P", "Discovery Method", 18),
        ("Q", "Discovery Date", 15),
        ("R", "Notes/Comments", 30),
    ]
    
    for col, header_text, width in headers:
        cell = ws[f'{col}6']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # EXAMPLE ROW (Row 7) - Gray italic
    example_data = [
        "DE-001",
        "Customer CRM System",
        "CRM_PROD",
        "Customers",
        "email_address",
        "String",
        "CAT-PII-D",
        "Email",
        "High",
        "Yes",
        "GDPR, FADP",
        "Yes",
        "Not Masked",
        "500000",
        "VP Sales",
        "Schema Review",
        "2024-11-15",
        "Primary customer contact email",
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=7, column=col_idx)
        cell.value = value
        apply_style(cell, styles["info_cell"])
    
    # DATA ENTRY ROWS (8-107: 100 rows)
    for row_idx in range(8, 108):
        for col_idx in range(1, 19):  # A-R = 18 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Apply data validations to appropriate columns
    for row in range(8, 108):
        validations['data_type'].add(ws[f'F{row}'])
        validations['data_category'].add(ws[f'G{row}'])
        validations['pii_type'].add(ws[f'H{row}'])
        validations['sensitivity_level'].add(ws[f'I{row}'])
        validations['yes_no'].add(ws[f'J{row}'])
        validations['yes_no_conditional'].add(ws[f'L{row}'])
        validations['masking_status'].add(ws[f'M{row}'])
        validations['discovery_method'].add(ws[f'P{row}'])
    
    # CHECKLIST SECTION (Starting Row 109)
    checklist_row = 109
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "SENSITIVE DATA INVENTORY CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    # Checklist headers
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Is sensitive data identified at field level?",
        "Are all production databases inventoried?",
        "Are non-production environments inventoried?",
        "Is PII specifically flagged?",
        "Are special category data (GDPR Art.9) identified?",
        "Is financial data (PCI scope) identified?",
        "Are credentials/secrets identified?",
        "Is data category assigned per field?",
        "Is sensitivity level assigned?",
        "Is masking requirement determined?",
        "Is current masking status documented?",
        "Are record counts estimated?",
        "Is data ownership assigned?",
        "Is discovery method documented?",
        "Are regulatory requirements mapped?",
        "Is inventory updated within 30 days of changes?",
        "Is automated discovery used where feasible?",
        "Are discovery results validated manually?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        ws.row_dimensions[checklist_row].height = 20
        checklist_row += 1
    
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    
    # Freeze panes
    ws.freeze_panes = "A7"


# ============================================================================
# SECTION 8: SHEET 5 - CLASSIFICATION MATRIX
# ============================================================================

def create_classification_matrix(ws, styles):
    """Create Classification Matrix sheet."""
    
    validations = create_base_validations(ws)
    
    # HEADER
    ws.merge_cells('A1:L1')
    header = ws['A1']
    header.value = "DATA CLASSIFICATION MATRIX"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 25
    
    # SUBTITLE
    ws.merge_cells('A2:L2')
    subtitle = ws['A2']
    subtitle.value = "Map data sensitivity to organizational classification scheme"
    apply_style(subtitle, styles["subheader"])
    ws.row_dimensions[2].height = 20
    
    # ASSESSMENT QUESTION
    ws.merge_cells('A3:L3')
    ws['A3'] = "Does your organization classify sensitive data according to an established taxonomy?"
    ws['A3'].font = Font(bold=True, size=11)
    ws.row_dimensions[3].height = 25
    
    # RESPONSE DROPDOWN
    ws['A4'] = "Response:"
    ws['A4'].font = Font(bold=True)
    apply_style(ws['B4'], styles["input_cell"])
    validations['yes_no_partial_planned_na'].add(ws['B4'])
    
    # CLASSIFICATION SUMMARY TABLE (Rows 6-11)
    ws.merge_cells('A6:L6')
    ws['A6'] = "CLASSIFICATION SUMMARY"
    ws['A6'].font = Font(bold=True, size=11)
    
    # Summary headers (Row 7)
    summary_headers = [
        ("A", "Organizational Class", 20),
        ("B", "Typical Sensitivity", 18),
        ("C", "Data Categories", 25),
        ("D", "Masking Requirement", 30),
        ("E", "Count of Fields", 15),
    ]
    
    for col, header_text, width in summary_headers:
        cell = ws[f'{col}7']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # Summary data (Rows 8-11)
    summary_data = [
        ("Restricted", "Critical", "CAT-HLT, CAT-BIO, CAT-GEN, CAT-FIN (PAN)", "Mandatory in ALL non-prod"),
        ("Confidential", "High", "CAT-PII-D, CAT-FIN, CAT-CRD", "Mandatory in non-prod"),
        ("Internal", "Medium", "CAT-PII-I, CAT-PRP, CAT-LOC", "Risk-based masking"),
        ("Public", "Low/None", "Non-sensitive data", "No masking required"),
    ]
    
    for row_idx, (org_class, sensitivity, categories, masking_req) in enumerate(summary_data, start=8):
        ws[f'A{row_idx}'] = org_class
        ws[f'B{row_idx}'] = sensitivity
        ws[f'C{row_idx}'] = categories
        ws[f'D{row_idx}'] = masking_req
        ws[f'E{row_idx}'] = '=COUNTIF(Classification_Matrix!F:F,"' + org_class + '")'
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row_idx}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row_idx].height = 20
    
    # CLASSIFICATION ASSIGNMENT TABLE (Starting Row 14)
    ws.merge_cells('A14:L14')
    ws['A14'] = "CLASSIFICATION ASSIGNMENT (100 rows)"
    ws['A14'].font = Font(bold=True, size=11)
    
    # Column headers (Row 15)
    assign_headers = [
        ("A", "Data Element ID", 15),
        ("B", "System Name", 25),
        ("C", "Field Name", 25),
        ("D", "Data Category", 15),
        ("E", "Sensitivity Level", 15),
        ("F", "Organizational Classification", 20),
        ("G", "Masking Requirement", 18),
        ("H", "Classification Rationale", 30),
        ("I", "Classified By", 20),
        ("J", "Classification Date", 15),
        ("K", "Last Review Date", 15),
        ("L", "Next Review Date", 15),
    ]
    
    for col, header_text, width in assign_headers:
        cell = ws[f'{col}15']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # DATA ENTRY ROWS (16-115: 100 rows)
    for row_idx in range(16, 116):
        for col_idx in range(1, 13):  # A-L = 12 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Apply data validations
    for row in range(16, 116):
        validations['org_classification'].add(ws[f'F{row}'])
    
    # CHECKLIST (Starting Row 117)
    checklist_row = 117
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "CLASSIFICATION CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Is organizational classification scheme defined?",
        "Are all sensitive fields classified?",
        "Is classification aligned with data categories?",
        "Are masking requirements derived from classification?",
        "Is classification rationale documented?",
        "Are classifiers identified per field?",
        "Is classification reviewed annually?",
        "Are regulatory overrides documented?",
        "Is reclassification process defined?",
        "Are classification changes tracked?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        ws.row_dimensions[checklist_row].height = 20
        checklist_row += 1
    
    ws.column_dimensions['B'].width = 45
    
    ws.freeze_panes = "A16"


# ============================================================================
# SECTION 9: SHEET 6 - REGULATORY MAPPING
# ============================================================================

def create_regulatory_mapping(ws, styles):
    """Create Regulatory Mapping sheet."""
    
    validations = create_base_validations(ws)
    
    # HEADER
    ws.merge_cells('A1:N1')
    header = ws['A1']
    header.value = "REGULATORY REQUIREMENT MAPPING"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 25
    
    # SUBTITLE
    ws.merge_cells('A2:N2')
    subtitle = ws['A2']
    subtitle.value = "Map data to applicable privacy and security regulations"
    apply_style(subtitle, styles["subheader"])
    ws.row_dimensions[2].height = 20
    
    # ASSESSMENT QUESTION
    ws.merge_cells('A3:N3')
    ws['A3'] = "Has your organization identified which data is subject to specific regulatory requirements?"
    ws['A3'].font = Font(bold=True, size=11)
    ws.row_dimensions[3].height = 25
    
    # RESPONSE DROPDOWN
    ws['A4'] = "Response:"
    ws['A4'].font = Font(bold=True)
    apply_style(ws['B4'], styles["input_cell"])
    validations['yes_no_partial_planned_na'].add(ws['B4'])
    
    # REGULATORY SUMMARY (Rows 6-13)
    ws.merge_cells('A6:E6')
    ws['A6'] = "REGULATORY SUMMARY"
    ws['A6'].font = Font(bold=True, size=11)
    
    reg_headers = [
        ("A", "Regulation", 20),
        ("B", "Applicability", 15),
        ("C", "Data Categories in Scope", 30),
        ("D", "Field Count", 12),
        ("E", "Masking Mandated?", 18),
    ]
    
    for col, header_text, width in reg_headers:
        cell = ws[f'{col}7']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    regulations = [
        ("GDPR (EU)", "Art. 32 Security", "Yes (Art. 32)"),
        ("FADP (Switzerland)", "Data Protection", "Yes"),
        ("HIPAA (US - if applicable)", "Protected Health Info", "Yes"),
        ("PCI-DSS (if processing cards)", "Cardholder Data", "Yes (Req 3.4)"),
        ("CCPA/CPRA (California)", "Consumer Privacy", "Recommended"),
        ("Other (specify)", "", ""),
    ]
    
    for row_idx, (regulation, scope, masking) in enumerate(regulations, start=8):
        ws[f'A{row_idx}'] = regulation
        ws[f'B{row_idx}'].value = ""
        apply_style(ws[f'B{row_idx}'], styles["input_cell"])
        validations['yes_no_partial_planned_na'].add(ws[f'B{row_idx}'])
        ws[f'C{row_idx}'] = scope
        ws[f'D{row_idx}'] = '=COUNTIF(Regulatory_Mapping!E:E,"Yes")'
        ws[f'E{row_idx}'] = masking
        ws.row_dimensions[row_idx].height = 20
    
    # DETAILED MAPPING TABLE (Starting Row 15)
    ws.merge_cells('A15:N15')
    ws['A15'] = "DETAILED REGULATORY MAPPING (100 rows)"
    ws['A15'].font = Font(bold=True, size=11)
    
    map_headers = [
        ("A", "Data Element ID", 15),
        ("B", "System Name", 25),
        ("C", "Field Name", 25),
        ("D", "Data Category", 15),
        ("E", "GDPR Applicable?", 15),
        ("F", "GDPR Article", 15),
        ("G", "FADP Applicable?", 15),
        ("H", "HIPAA Applicable?", 15),
        ("I", "PCI-DSS Applicable?", 15),
        ("J", "Other Regulations", 20),
        ("K", "Masking Mandated by Reg?", 18),
        ("L", "Regulatory Reference", 25),
        ("M", "Compliance Deadline", 15),
        ("N", "Notes", 30),
    ]
    
    for col, header_text, width in map_headers:
        cell = ws[f'{col}16']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # DATA ENTRY ROWS (17-116: 100 rows)
    for row_idx in range(17, 117):
        for col_idx in range(1, 15):  # A-N = 14 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Apply data validations
    for row in range(17, 117):
        validations['yes_no_unknown'].add(ws[f'E{row}'])
        validations['gdpr_article'].add(ws[f'F{row}'])
        validations['yes_no_unknown'].add(ws[f'G{row}'])
        validations['yes_no_unknown'].add(ws[f'H{row}'])
        validations['yes_no_unknown'].add(ws[f'I{row}'])
        validations['yes_no_partial'].add(ws[f'K{row}'])
    
    # CHECKLIST (Starting Row 118)
    checklist_row = 118
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "REGULATORY COMPLIANCE CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Are applicable regulations identified?",
        "Is GDPR applicability determined?",
        "Is FADP applicability determined?",
        "Is HIPAA applicability determined (if US)?",
        "Is PCI-DSS applicability determined?",
        "Are GDPR special categories (Art.9) identified?",
        "Is PCI-DSS PAN data identified?",
        "Are masking requirements per regulation documented?",
        "Are regulatory references cited?",
        "Are compliance deadlines tracked?",
        "Is DPO consulted for GDPR scope?",
        "Are cross-border data transfers considered?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        ws.row_dimensions[checklist_row].height = 20
        checklist_row += 1
    
    ws.column_dimensions['B'].width = 45
    
    ws.freeze_panes = "A17"


# ============================================================================
# SECTION 10: SHEET 7 - DATA OWNER ASSIGNMENT
# ============================================================================

def create_data_owner_assignment(ws, styles):
    """Create Data Owner Assignment sheet."""
    
    validations = create_base_validations(ws)
    
    # HEADER
    ws.merge_cells('A1:N1')
    header = ws['A1']
    header.value = "DATA OWNERSHIP ASSIGNMENT"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 25
    
    # SUBTITLE
    ws.merge_cells('A2:N2')
    subtitle = ws['A2']
    subtitle.value = "Assign data owners per system and data category (ISMS-POL-A.8.11-S3)"
    apply_style(subtitle, styles["subheader"])
    ws.row_dimensions[2].height = 20
    
    # ASSESSMENT QUESTION
    ws.merge_cells('A3:N3')
    ws['A3'] = "Has your organization assigned data owners for all sensitive data categories?"
    ws['A3'].font = Font(bold=True, size=11)
    ws.row_dimensions[3].height = 25
    
    # RESPONSE DROPDOWN
    ws['A4'] = "Response:"
    ws['A4'].font = Font(bold=True)
    apply_style(ws['B4'], styles["input_cell"])
    validations['yes_no_partial_planned_na'].add(ws['B4'])
    
    # OWNERSHIP SUMMARY BY CATEGORY (Rows 6-17)
    ws.merge_cells('A6:F6')
    ws['A6'] = "OWNERSHIP SUMMARY BY DATA CATEGORY"
    ws['A6'].font = Font(bold=True, size=11)
    
    owner_headers = [
        ("A", "Data Category", 20),
        ("B", "Data Owner (Role)", 20),
        ("C", "Data Owner (Name)", 20),
        ("D", "Backup Owner", 20),
        ("E", "Assignment Date", 15),
        ("F", "Last Review", 15),
    ]
    
    for col, header_text, width in owner_headers:
        cell = ws[f'{col}7']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    data_categories = [
        "CAT-PII-D (Direct PII)",
        "CAT-PII-I (Indirect PII)",
        "CAT-FIN (Financial)",
        "CAT-HLT (Health)",
        "CAT-CRD (Credentials)",
        "CAT-PRP (Proprietary)",
        "CAT-LOC (Location)",
        "CAT-BIO (Biometric)",
        "CAT-GEN (Genetic)",
        "CAT-CHD (Child Data)",
    ]
    
    for row_idx, category in enumerate(data_categories, start=8):
        ws[f'A{row_idx}'] = category
        for col in ['B', 'C', 'D', 'E', 'F']:
            apply_style(ws[f'{col}{row_idx}'], styles["input_cell"])
        ws.row_dimensions[row_idx].height = 20
    
    # OWNERSHIP BY SYSTEM (Starting Row 19)
    ws.merge_cells('A19:N19')
    ws['A19'] = "OWNERSHIP BY SYSTEM (50 rows)"
    ws['A19'].font = Font(bold=True, size=11)
    
    system_headers = [
        ("A", "System Name", 25),
        ("B", "System Owner", 20),
        ("C", "Primary Data Owner", 20),
        ("D", "Data Steward", 20),
        ("E", "Business Owner", 20),
        ("F", "Data Categories in System", 25),
        ("G", "Ownership Documented?", 15),
        ("H", "RACI Matrix Reference", 20),
        ("I", "Assignment Date", 15),
        ("J", "Last Review Date", 15),
        ("K", "Next Review Date", 15),
        ("L", "Approval Status", 15),
        ("M", "Approver", 20),
        ("N", "Notes", 30),
    ]
    
    for col, header_text, width in system_headers:
        cell = ws[f'{col}20']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # DATA ENTRY ROWS (21-70: 50 rows)
    for row_idx in range(21, 71):
        for col_idx in range(1, 15):  # A-N = 14 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Apply data validations
    for row in range(21, 71):
        validations['yes_no_partial'].add(ws[f'G{row}'])
    
    # DATA OWNER RESPONSIBILITIES REFERENCE (Starting Row 72)
    ws.merge_cells('A72:C72')
    ws['A72'] = "DATA OWNER RESPONSIBILITIES (Reference)"
    ws['A72'].font = Font(bold=True, size=11)
    
    ws['A73'] = "Responsibility"
    ws['B73'] = "Requirement Level"
    ws['C73'] = "Policy Reference"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f'{col}73'], styles["column_header"])
    
    responsibilities = [
        ("Approve data classification", "SHALL", "REQ-CLS-031"),
        ("Approve masking techniques for their data", "SHALL", "REQ-CLS-031"),
        ("Review classification annually", "SHALL", "REQ-CLS-032"),
        ("Approve access to unmasked data", "SHALL", "REQ-ENV-041"),
        ("Review masking effectiveness", "SHOULD", "REQ-TST-020"),
        ("Participate in risk assessments", "SHOULD", "REQ-GOV-010"),
    ]
    
    for row_idx, (responsibility, level, policy_ref) in enumerate(responsibilities, start=74):
        ws[f'A{row_idx}'] = responsibility
        ws[f'B{row_idx}'] = level
        ws[f'C{row_idx}'] = policy_ref
        ws.row_dimensions[row_idx].height = 20
    
    # CHECKLIST (Starting Row 81)
    checklist_row = 81
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "OWNERSHIP CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Are data owners assigned per category?",
        "Are data owners assigned per system?",
        "Are data stewards identified?",
        "Are backup owners assigned?",
        "Is ownership documented in RACI matrix?",
        "Have data owners acknowledged responsibilities?",
        "Are ownership changes tracked?",
        "Is ownership reviewed annually?",
        "Are data owners trained on masking requirements?",
        "Do data owners approve masking techniques?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        ws.row_dimensions[checklist_row].height = 20
        checklist_row += 1
    
    ws.column_dimensions['B'].width = 45
    
    ws.freeze_panes = "A21"


# ============================================================================
# SECTION 11: SHEET 8 - MASKING PRIORITY MATRIX
# ============================================================================

def create_masking_priority_matrix(ws, styles):
    """Create Masking Priority Matrix sheet with risk-based scoring."""
    
    validations = create_base_validations(ws)
    
    # HEADER
    ws.merge_cells('A1:P1')
    header = ws['A1']
    header.value = "MASKING PRIORITY & RISK ASSESSMENT"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 25
    
    # SUBTITLE
    ws.merge_cells('A2:P2')
    subtitle = ws['A2']
    subtitle.value = "Prioritize masking efforts based on risk, compliance, and business impact"
    apply_style(subtitle, styles["subheader"])
    ws.row_dimensions[2].height = 20
    
    # ASSESSMENT QUESTION
    ws.merge_cells('A3:P3')
    ws['A3'] = "Has your organization prioritized masking implementation based on risk assessment?"
    ws['A3'].font = Font(bold=True, size=11)
    ws.row_dimensions[3].height = 25
    
    # RESPONSE DROPDOWN
    ws['A4'] = "Response:"
    ws['A4'].font = Font(bold=True)
    apply_style(ws['B4'], styles["input_cell"])
    validations['yes_no_partial_planned_na'].add(ws['B4'])
    
    # PRIORITY CALCULATION FORMULA (Rows 6-10)
    ws.merge_cells('A6:P6')
    ws['A6'] = "PRIORITY SCORE FORMULA: (Sensitivity × 3) + (Exposure Risk × 2) + (Regulatory Weight × 2) + (Volume Factor × 1)"
    ws['A6'].font = Font(bold=True, size=10)
    ws['A6'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[6].height = 25
    
    ws['A8'] = "Priority Tiers:"
    ws['A8'].font = Font(bold=True)
    tiers = [
        "P1 (Critical): Score ≥ 35 — Immediate action required",
        "P2 (High): Score 25-34 — Implement within 3 months",
        "P3 (Medium): Score 15-24 — Implement within 6 months",
        "P4 (Low): Score < 15 — Risk-based timeline",
    ]
    for idx, tier in enumerate(tiers, start=9):
        ws[f'A{idx}'] = tier
        ws.row_dimensions[idx].height = 15
    
    # PRIORITY ASSESSMENT TABLE (Starting Row 13)
    ws.merge_cells('A13:P13')
    ws['A13'] = "PRIORITY ASSESSMENT (100 rows)"
    ws['A13'].font = Font(bold=True, size=11)
    
    priority_headers = [
        ("A", "Data Element ID", 15),
        ("B", "System Name", 25),
        ("C", "Field Name", 25),
        ("D", "Data Category", 15),
        ("E", "Sensitivity Score (1-5)", 12),
        ("F", "Exposure Risk (1-5)", 15),
        ("G", "Regulatory Score (1-5)", 15),
        ("H", "Volume Score (1-5)", 12),
        ("I", "Total Priority Score", 15),
        ("J", "Priority Tier", 12),
        ("K", "Current Masking Status", 18),
        ("L", "Target Implementation Date", 18),
        ("M", "Assigned To", 20),
        ("N", "Implementation Status", 18),
        ("O", "Blocking Issues", 25),
        ("P", "Notes", 30),
    ]
    
    for col, header_text, width in priority_headers:
        cell = ws[f'{col}14']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # DATA ENTRY ROWS (15-114: 100 rows)
    for row_idx in range(15, 115):
        for col_idx in range(1, 17):  # A-P = 16 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Apply data validations
    for row in range(15, 115):
        validations['exposure_risk'].add(ws[f'F{row}'])
        validations['implementation_status'].add(ws[f'N{row}'])
        
        # Formulas for calculated fields
        # Total Priority Score: (E×3)+(F×2)+(G×2)+(H×1) - requires manual calculation or VBA
        ws[f'I{row}'] = ""  # User enters or calculates externally
        # Priority Tier based on score - requires manual entry
        ws[f'J{row}'] = ""
    
    # PRIORITY SUMMARY DASHBOARD (Starting Row 116)
    ws.merge_cells('A116:G116')
    ws['A116'] = "PRIORITY TIER SUMMARY"
    ws['A116'].font = Font(bold=True, size=11)
    
    summary_headers = [
        ("A", "Priority Tier", 15),
        ("B", "Count", 10),
        ("C", "% of Total", 12),
        ("D", "Avg Score", 12),
        ("E", "Complete", 12),
        ("F", "In Progress", 12),
        ("G", "Not Started", 12),
    ]
    
    for col, header_text, width in summary_headers:
        cell = ws[f'{col}117']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    priority_tiers = ["P1 (Critical)", "P2 (High)", "P3 (Medium)", "P4 (Low)", "TOTAL"]
    
    for row_idx, tier in enumerate(priority_tiers, start=118):
        ws[f'A{row_idx}'] = tier
        # Formulas would go here - simplified for generator
        ws[f'B{row_idx}'] = ""
        ws[f'C{row_idx}'] = ""
        ws[f'D{row_idx}'] = ""
        ws[f'E{row_idx}'] = ""
        ws[f'F{row_idx}'] = ""
        ws[f'G{row_idx}'] = ""
        ws.row_dimensions[row_idx].height = 20
    
    # CHECKLIST (Starting Row 124)
    checklist_row = 124
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "PRIORITY CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Is risk-based prioritization performed?",
        "Are sensitivity levels factored into priority?",
        "Is exposure risk assessed per field?",
        "Are regulatory mandates weighted?",
        "Is data volume considered?",
        "Are P1 items addressed immediately?",
        "Are target dates assigned per priority?",
        "Are responsible parties assigned?",
        "Is implementation status tracked?",
        "Are blocking issues documented?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        ws.row_dimensions[checklist_row].height = 20
        checklist_row += 1
    
    ws.column_dimensions['B'].width = 45
    
    ws.freeze_panes = "A15"


# ============================================================================
# SECTION 12: SHEET 9 - GAP ANALYSIS
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet."""
    
    validations = create_base_validations(ws)
    
    # HEADER
    ws.merge_cells('A1:O1')
    header = ws['A1']
    header.value = "DATA INVENTORY & CLASSIFICATION GAP ANALYSIS"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 25
    
    # SUBTITLE
    ws.merge_cells('A2:O2')
    subtitle = ws['A2']
    subtitle.value = "Identify missing inventory, classification, ownership, or masking"
    apply_style(subtitle, styles["subheader"])
    ws.row_dimensions[2].height = 20
    
    # GAP SUMMARY TABLE (Rows 4-12)
    ws.merge_cells('A4:F4')
    ws['A4'] = "GAP SUMMARY"
    ws['A4'].font = Font(bold=True, size=11)
    
    gap_summary_headers = [
        ("A", "Gap Category", 25),
        ("B", "Count", 10),
        ("C", "% of Total", 12),
        ("D", "Risk Level", 12),
        ("E", "Remediation Owner", 20),
        ("F", "Target Date", 15),
    ]
    
    for col, header_text, width in gap_summary_headers:
        cell = ws[f'{col}5']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    gap_categories = [
        "Systems not inventoried",
        "Systems with unknown sensitive data",
        "Fields not classified",
        "Fields without data owner",
        "Regulatory mapping incomplete",
        "Masking required but not implemented",
        "Masking status unknown",
    ]
    
    for row_idx, category in enumerate(gap_categories, start=6):
        ws[f'A{row_idx}'] = category
        apply_style(ws[f'B{row_idx}'], styles["input_cell"])
        apply_style(ws[f'C{row_idx}'], styles["input_cell"])
        apply_style(ws[f'D{row_idx}'], styles["input_cell"])
        validations['risk_level'].add(ws[f'D{row_idx}'])
        apply_style(ws[f'E{row_idx}'], styles["input_cell"])
        apply_style(ws[f'F{row_idx}'], styles["input_cell"])
        ws.row_dimensions[row_idx].height = 20
    
    # DETAILED GAP REGISTER (Starting Row 14)
    ws.merge_cells('A14:O14')
    ws['A14'] = "DETAILED GAP REGISTER (60 rows)"
    ws['A14'].font = Font(bold=True, size=11)
    
    gap_headers = [
        ("A", "Gap ID", 12),
        ("B", "Gap Category", 20),
        ("C", "System Name", 25),
        ("D", "Field/Data Element", 25),
        ("E", "Gap Description", 35),
        ("F", "Risk Level", 12),
        ("G", "Impact if Not Remediated", 30),
        ("H", "Root Cause", 25),
        ("I", "Remediation Action", 30),
        ("J", "Remediation Owner", 20),
        ("K", "Target Completion Date", 15),
        ("L", "Status", 15),
        ("M", "Actual Completion Date", 15),
        ("N", "Verification Method", 20),
        ("O", "Notes", 30),
    ]
    
    for col, header_text, width in gap_headers:
        cell = ws[f'{col}15']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # DATA ENTRY ROWS (16-75: 60 rows)
    for row_idx in range(16, 76):
        for col_idx in range(1, 16):  # A-O = 15 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Apply data validations
    for row in range(16, 76):
        validations['gap_category'].add(ws[f'B{row}'])
        validations['risk_level'].add(ws[f'F{row}'])
        validations['gap_status'].add(ws[f'L{row}'])
    
    # CHECKLIST (Starting Row 77)
    checklist_row = 77
    ws.merge_cells(f'A{checklist_row}:E{checklist_row}')
    ws[f'A{checklist_row}'] = "GAP ANALYSIS CHECKLIST"
    ws[f'A{checklist_row}'].font = Font(bold=True, size=11)
    checklist_row += 1
    
    ws[f'A{checklist_row}'] = "#"
    ws[f'B{checklist_row}'] = "Checklist Item"
    ws[f'C{checklist_row}'] = "Status"
    ws[f'D{checklist_row}'] = "Evidence"
    ws[f'E{checklist_row}'] = "Notes"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{checklist_row}'], styles["column_header"])
    checklist_row += 1
    
    checklist_items = [
        "Is gap analysis performed quarterly?",
        "Are missing inventories identified?",
        "Are unclassified systems identified?",
        "Are fields without owners identified?",
        "Are regulatory gaps documented?",
        "Are masking gaps prioritized?",
        "Are root causes analyzed?",
        "Are remediation actions defined?",
        "Are remediation owners assigned?",
        "Are target dates realistic?",
        "Is gap closure tracked?",
        "Are accepted risks documented?",
    ]
    
    for idx, item in enumerate(checklist_items, start=1):
        ws[f'A{checklist_row}'] = idx
        ws[f'B{checklist_row}'] = item
        ws[f'B{checklist_row}'].alignment = Alignment(wrap_text=True)
        apply_style(ws[f'C{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'D{checklist_row}'], styles["input_cell"])
        apply_style(ws[f'E{checklist_row}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'C{checklist_row}'])
        ws.row_dimensions[checklist_row].height = 20
        checklist_row += 1
    
    ws.column_dimensions['B'].width = 45
    
    ws.freeze_panes = "A16"


# ============================================================================
# SECTION 13: SHEET 10 - EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register with 100-row template."""
    
    validations = create_base_validations(ws)
    
    # HEADER
    ws.merge_cells('A1:P1')
    header = ws['A1']
    header.value = "EVIDENCE REGISTER"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 25
    
    # SUBTITLE
    ws.merge_cells('A2:P2')
    subtitle = ws['A2']
    subtitle.value = "Audit trail of supporting documentation (100 entry template)"
    apply_style(subtitle, styles["subheader"])
    ws.row_dimensions[2].height = 20
    
    # COLUMN HEADERS (Row 4)
    evidence_headers = [
        ("A", "Evidence ID", 12),
        ("B", "Evidence Type", 20),
        ("C", "Evidence Title", 30),
        ("D", "Related System(s)", 25),
        ("E", "Related Requirement(s)", 20),
        ("F", "Related Sheet/Assessment", 20),
        ("G", "Evidence Location", 30),
        ("H", "Evidence Date", 15),
        ("I", "Evidence Owner", 20),
        ("J", "Retention Period", 15),
        ("K", "Review Frequency", 15),
        ("L", "Last Reviewed Date", 15),
        ("M", "Next Review Date", 15),
        ("N", "Confidentiality Level", 15),
        ("O", "Access Restrictions", 25),
        ("P", "Notes", 30),
    ]
    
    for col, header_text, width in evidence_headers:
        cell = ws[f'{col}4']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # DATA ENTRY ROWS (5-104: 100 rows)
    for row_idx in range(5, 105):
        for col_idx in range(1, 17):  # A-P = 16 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Apply data validations
    for row in range(5, 105):
        validations['evidence_type'].add(ws[f'B{row}'])
        validations['review_frequency'].add(ws[f'K{row}'])
        validations['confidentiality_level'].add(ws[f'N{row}'])
    
    # EVIDENCE TYPES REFERENCE (Starting Row 106)
    ws.merge_cells('A106:C106')
    ws['A106'] = "EVIDENCE TYPE DEFINITIONS"
    ws['A106'].font = Font(bold=True, size=11)
    
    ws['A107'] = "Evidence Type"
    ws['B107'] = "Description"
    ws['C107'] = "Typical Retention"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f'{col}107'], styles["column_header"])
    
    evidence_types = [
        ("Data Discovery Report", "Automated or manual data discovery results", "3 years"),
        ("Schema Documentation", "Database schemas, data dictionaries", "Current + 1 year"),
        ("DPIA", "Data Protection Impact Assessments", "3 years post-processing"),
        ("Data Flow Diagram", "How data moves between systems", "Current + 2 years"),
        ("Classification Review", "Annual classification review records", "3 years"),
        ("RACI Matrix", "Responsibility assignments", "Current + 1 year"),
        ("Meeting Minutes", "Classification meetings, approvals", "3 years"),
        ("Approval Email", "Data owner approvals, sign-offs", "3 years"),
        ("Tool Report", "Reports from discovery/masking tools", "1 year"),
    ]
    
    for row_idx, (ev_type, description, retention) in enumerate(evidence_types, start=108):
        ws[f'A{row_idx}'] = ev_type
        ws[f'B{row_idx}'] = description
        ws[f'C{row_idx}'] = retention
        ws.row_dimensions[row_idx].height = 20
    
    ws.column_dimensions['B'].width = 40
    
    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 14: SHEET 11 - SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with KPIs and executive summary."""
    
    validations = create_base_validations(ws)
    
    # HEADER
    ws.merge_cells('A1:H1')
    header = ws['A1']
    header.value = "DATA INVENTORY & CLASSIFICATION DASHBOARD"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 25
    
    # SUBTITLE
    ws.merge_cells('A2:H2')
    subtitle = ws['A2']
    subtitle.value = "Executive summary with key metrics and compliance status"
    apply_style(subtitle, styles["subheader"])
    ws.row_dimensions[2].height = 20
    
    # COMPLETION STATUS SECTION (Rows 4-9)
    ws.merge_cells('A4:G4')
    ws['A4'] = "COMPLETION STATUS"
    ws['A4'].font = Font(bold=True, size=11)
    
    completion_headers = [
        ("A", "Assessment Area", 25),
        ("B", "Total Items", 12),
        ("C", "Complete", 12),
        ("D", "Partial", 12),
        ("E", "Missing", 12),
        ("F", "Planned", 12),
        ("G", "% Complete", 12),
    ]
    
    for col, header_text, width in completion_headers:
        cell = ws[f'{col}5']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    completion_areas = [
        "Systems Inventoried",
        "Sensitive Data Fields",
        "Classification Applied",
        "Regulatory Mapping",
        "Data Ownership",
    ]
    
    for row_idx, area in enumerate(completion_areas, start=6):
        ws[f'A{row_idx}'] = area
        # Formulas would go here - simplified for generator
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{row_idx}'] = ""
        ws.row_dimensions[row_idx].height = 20
    
    # DATA CATEGORY COVERAGE (Rows 11-23)
    ws.merge_cells('A11:G11')
    ws['A11'] = "DATA CATEGORY COVERAGE"
    ws['A11'].font = Font(bold=True, size=11)
    
    category_headers = [
        ("A", "Data Category", 20),
        ("B", "Fields Identified", 15),
        ("C", "% of Total", 12),
        ("D", "Classified", 12),
        ("E", "Owner Assigned", 15),
        ("F", "Masking Required", 15),
        ("G", "Masking Implemented", 18),
    ]
    
    for col, header_text, width in category_headers:
        cell = ws[f'{col}12']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    data_categories_list = [
        "CAT-PII-D",
        "CAT-PII-I",
        "CAT-FIN",
        "CAT-HLT",
        "CAT-CRD",
        "CAT-PRP",
        "CAT-LOC",
        "CAT-BIO",
        "CAT-GEN",
        "CAT-CHD",
        "TOTAL",
    ]
    
    for row_idx, category in enumerate(data_categories_list, start=13):
        ws[f'A{row_idx}'] = category
        # Formulas would go here
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{row_idx}'] = ""
        ws.row_dimensions[row_idx].height = 20
    
    # KEY PERFORMANCE INDICATORS (Rows 25-36)
    ws.merge_cells('A25:E25')
    ws['A25'] = "KEY PERFORMANCE INDICATORS"
    ws['A25'].font = Font(bold=True, size=11)
    
    kpi_headers = [
        ("A", "KPI", 35),
        ("B", "Current Value", 15),
        ("C", "Target", 12),
        ("D", "Status", 12),
        ("E", "Trend", 12),
    ]
    
    for col, header_text, width in kpi_headers:
        cell = ws[f'{col}26']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    kpis = [
        ("% Systems Inventoried", "100%"),
        ("% Sensitive Fields Classified", "100%"),
        ("% Data Owners Assigned", "100%"),
        ("% Regulatory Mapping Complete", "100%"),
        ("% Masking Requirements Defined", "100%"),
        ("% P1 Items Complete", "100%"),
        ("Mean Time to Classify (days)", "<30 days"),
        ("Inventory Accuracy (last audit)", ">95%"),
        ("Open Critical Gaps", "0"),
    ]
    
    for row_idx, (kpi, target) in enumerate(kpis, start=27):
        ws[f'A{row_idx}'] = kpi
        ws[f'B{row_idx}'] = ""
        ws[f'C{row_idx}'] = target
        ws[f'D{row_idx}'] = ""
        ws[f'E{row_idx}'] = ""
        ws.row_dimensions[row_idx].height = 20
    
    # ASSESSMENT SIGN-OFF (Rows 38-46)
    ws.merge_cells('A38:E38')
    ws['A38'] = "ASSESSMENT SIGN-OFF"
    ws['A38'].font = Font(bold=True, size=11)
    
    signoff_headers = [
        ("A", "Role", 30),
        ("B", "Name", 25),
        ("C", "Signature", 25),
        ("D", "Date", 15),
        ("E", "Comments", 30),
    ]
    
    for col, header_text, width in signoff_headers:
        cell = ws[f'{col}39']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    signoff_roles = [
        "Data Governance Lead",
        "Chief Data Officer (CDO)",
        "Data Protection Officer (DPO)",
        "Chief Information Security Officer (CISO)",
        "Legal/Compliance Officer",
    ]
    
    for row_idx, role in enumerate(signoff_roles, start=40):
        ws[f'A{row_idx}'] = role
        for col in ['B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row_idx}'], styles["input_cell"])
        ws.row_dimensions[row_idx].height = 20
    
    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 15: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    try:
        logger.info("=" * 78)
        logger.info("%s - %s Generator", WORKBOOK_ID, ASSESSMENT_AREA)
        logger.info("ISO/IEC 27001:2022 Control A.8.11: Data Masking")
        logger.info("=" * 78)

        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/11] Creating Instructions & Legend...")
        create_instructions_legend(wb["Instructions_Legend"], styles)

        logger.info("[2/11] Creating System Inventory...")
        create_system_inventory(wb["System_Inventory"], styles)

        logger.info("[3/11] Creating Data Category Reference...")
        create_data_category_reference(wb["Data_Category_Reference"], styles)

        logger.info("[4/11] Creating Sensitive Data Inventory...")
        create_sensitive_data_inventory(wb["Sensitive_Data_Inventory"], styles)

        logger.info("[5/11] Creating Classification Matrix...")
        create_classification_matrix(wb["Classification_Matrix"], styles)

        logger.info("[6/11] Creating Regulatory Mapping...")
        create_regulatory_mapping(wb["Regulatory_Mapping"], styles)

        logger.info("[7/11] Creating Data Owner Assignment...")
        create_data_owner_assignment(wb["Data_Owner_Assignment"], styles)

        logger.info("[8/11] Creating Masking Priority Matrix...")
        create_masking_priority_matrix(wb["Masking_Priority_Matrix"], styles)

        logger.info("[9/11] Creating Gap Analysis...")
        create_gap_analysis(wb["Gap_Analysis"], styles)

        logger.info("[10/11] Creating Evidence Register...")
        create_evidence_register(wb["Evidence_Register"], styles)

        logger.info("[11/11] Creating Summary Dashboard...")
        create_summary_dashboard(wb["Summary_Dashboard"], styles)

        # Save workbook
        filename = f"ISMS-IMP-A.8.11.1_Data_Inventory_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)

        logger.info("SUCCESS: %s", filename)
        logger.info("Workbook Structure: 11 sheets including Instructions, System Inventory, Evidence Register")
        logger.info("=" * 78)
        return 0
    except Exception as e:
        logger.error("Failed to generate workbook: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
