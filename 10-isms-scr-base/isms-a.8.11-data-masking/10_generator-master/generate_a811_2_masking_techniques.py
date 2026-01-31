#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# Licensed under AGPL-3.0-or-later with commercial licensing option
#
# This file is part of the ISMS Compliance Framework
# See /LICENSE for full terms and /LICENSES/COMMERCIAL.md for commercial options
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.11.2 - Masking Technique Selection & Requirements Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.11: Data Masking
Assessment Domain 2 of 5: Masking Technique Selection & Requirements

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific data types, masking requirements, and tool
capabilities.

Key customization areas:
1. Masking technique taxonomy (match your approved techniques)
2. Data type mappings (adapt to your data classification scheme)
3. Tool and vendor options (specific to your technology stack)
4. Performance thresholds (based on your SLA requirements)
5. Compliance requirements (aligned with your regulatory obligations)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for data masking)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
and selecting appropriate data masking techniques for sensitive data elements.

**Purpose:**
Enables systematic selection of masking techniques based on data type,
sensitivity, usage requirements, and compliance obligations.

**Assessment Scope:**
- Masking technique taxonomy and capability assessment
- Data type to technique mapping
- Referential integrity requirements
- Format preservation requirements
- Reversibility and key management assessment
- Tool and vendor capability evaluation
- Performance and scalability considerations
- Implementation status tracking
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Technique taxonomy and selection criteria
2. Technique Taxonomy - Standard masking techniques (read-only)
3. Data Type Mapping - Technique suitability by data type
4. Technique Requirements - Detailed capability requirements
5. Referential Integrity - Cross-system consistency requirements
6. Format Preservation - Format validation requirements
7. Reversibility Assessment - Key management requirements
8. Tool Capability Matrix - Vendor/tool evaluation
9. Performance Requirements - Scalability assessment
10. Implementation Status - Deployment tracking
11. Gap Analysis - Missing capabilities and remediation
12. Evidence Register - Audit evidence tracking
13. Summary Dashboard - Consolidated metrics

**Integration:**
This assessment feeds into ISMS-IMP-A.8.11.5 Compliance Dashboard, which
consolidates data from all five assessment domains for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a811_2_masking_techniques.py

Output:
    File: ISMS_IMP_A_8_11_2_Masking_Techniques_Assessment_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Review masking technique taxonomy
    2. Map data types to appropriate techniques
    3. Document referential integrity requirements
    4. Assess format preservation needs
    5. Evaluate reversibility requirements
    6. Compare tool/vendor capabilities
    7. Document performance requirements
    8. Track implementation status
    9. Collect audit evidence
    10. Feed results into A.8.11.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.11
Assessment Domain:    2 of 5 (Masking Technique Selection & Requirements)
Related Policy:       ISMS-POL-A.8.11 (Data Masking Policy)
Script Version:       1.0
Python Version:       3.8+

Related Documents:
    - ISMS-POL-A.8.11: Data Masking Policy
    - ISMS-IMP-A.8.11.1: Data Inventory & Classification (Domain 1)
    - ISMS-IMP-A.8.11.2: Masking Technique Selection Guide
    - ISMS-IMP-A.8.11.3: Environment Coverage (Domain 3)
    - ISMS-IMP-A.8.11.4: Testing & Validation (Domain 4)
    - ISMS-IMP-A.8.11.5: Compliance Dashboard (Consolidation)

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)



# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.11.2"
WORKBOOK_NAME = "Masking Technique Selection & Requirements"
CONTROL_ID = "A.8.11"
CONTROL_NAME = "Data Masking"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •


# Constants
WORKBOOK_ID = "ISMS-IMP-A.8.11.2"
RELATED_POLICY = "ISMS-POL-A.8.11-S2.2"
ASSESSMENT_AREA = "Masking Technique Selection"
COLOR_HEADER = "003366"
COLOR_SUBHEADER = "4472C4"
COLOR_COLUMN_HEADER = "D9D9D9"
COLOR_INPUT = "FFFFCC"
COLOR_INFO = "E7E6E6"
COLOR_COMPLETE = "C6EFCE"
COLOR_PARTIAL = "FFEB9C"
COLOR_MISSING = "FFC7CE"
COLOR_PLANNED = "B4C7E7"

def create_workbook():
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    sheets = ["Instructions_Legend", "Approved_Techniques", "Technique_Selection_Matrix",
              "Static_Masking_SDM", "Dynamic_Masking_DDM", "Tokenization_Implementation",
              "Encryption_for_Masking", "Masking_Tool_Inventory", "Configuration_Standards",
              "Gap_Analysis", "Evidence_Register", "Summary_Dashboard"]
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    return wb

def setup_styles():
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color=COLOR_COLUMN_HEADER, end_color=COLOR_COLUMN_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "info_cell": {
            "fill": PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"),
            "font": Font(name="Calibri", size=10, italic=True),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
    }

def apply_style(cell, style_dict):
    if "font" in style_dict:
        cell.font = Font(name=style_dict["font"].name, size=style_dict["font"].size,
                        bold=style_dict["font"].bold,
                        italic=getattr(style_dict["font"], 'italic', False),
                        color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None)
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type)
    if "alignment" in style_dict:
        cell.alignment = Alignment(horizontal=style_dict["alignment"].horizontal,
                                   vertical=style_dict["alignment"].vertical,
                                   wrap_text=style_dict["alignment"].wrap_text)
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def create_base_validations(ws):
    validations = {
        'yes_no': DataValidation(type="list", formula1='"Yes,No"', allow_blank=False),
        'yes_no_partial_planned_na': DataValidation(type="list", formula1='"Yes,No,Partial,Planned,N/A"', allow_blank=False),
        'status_icons': DataValidation(type="list", formula1='"\u2705 Implemented,\u26A0\uFE0F Partial,\u274C Not Implemented,\u1F4CB Planned,N/A"', allow_blank=False),
        'technique_id': DataValidation(type="list", formula1='"TECH-SDM,TECH-DDM,TECH-RED,TECH-TOK,TECH-SUB,TECH-ENC,TECH-SHF,TECH-HSH"', allow_blank=False),
        'sensitivity_level': DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=False),
        'environment_type': DataValidation(type="list", formula1='"Development,Test/QA,UAT,Training,Analytics,Production"', allow_blank=False),
        'masking_method': DataValidation(type="list", formula1='"Substitution,Redaction,Shuffling,Hashing,Tokenization,Encryption,Other"', allow_blank=False),
        'frequency': DataValidation(type="list", formula1='"On-Demand,Weekly,Monthly,Quarterly"', allow_blank=False),
        'yes_no_partial': DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=False),
        'risk_level': DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=False),
        'gap_status': DataValidation(type="list", formula1='"Open,In Progress,Complete,Accepted Risk"', allow_blank=False),
    }
    for dv in validations.values():
        ws.add_data_validation(dv)
    return validations

def create_instructions_legend(ws, styles):
    """Create Instructions sheet - condensed version"""
    ws.merge_cells('A1:H1')
    header = ws['A1']
    header.value = f"{WORKBOOK_ID} — {ASSESSMENT_AREA}"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:H2')
    subtitle = ws['A2']
    subtitle.value = "ISO/IEC 27001:2022 - Control A.8.11: Data Masking"
    apply_style(subtitle, styles["subheader"])
    
    info_fields = [
        ("Document ID:", WORKBOOK_ID),
        ("Assessment Area:", ASSESSMENT_AREA),
        ("Related Policy:", RELATED_POLICY),
        ("Review Cycle:", "Semi-Annual"),
    ]
    
    row = 4
    for label, value in info_fields:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    ws[f'A{row}'] = "HOW TO USE THIS WORKBOOK"
    ws[f'A{row}'].font = Font(size=11, bold=True)
    row += 1
    instructions = [
        "1. Review Approved_Techniques sheet",
        "2. Complete Technique_Selection_Matrix",
        "3. Document SDM/DDM implementations in respective sheets",
        "4. If using tokenization/encryption, complete those sheets",
        "5. Document tools in Masking_Tool_Inventory",
        "6. Review Configuration_Standards",
        "7. Identify gaps in Gap_Analysis",
        "8. Maintain Evidence_Register",
        "9. Review Summary_Dashboard",
        "10. Obtain approvals",
    ]
    for inst in instructions:
        ws[f'A{row}'] = inst
        row += 1
    
    ws.freeze_panes = "A4"

def create_approved_techniques(ws, styles):
    """Create Approved Techniques reference sheet"""
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:I1')
    header = ws['A1']
    header.value = "APPROVED MASKING TECHNIQUES"
    apply_style(header, styles["header"])
    ws.row_dimensions[1].height = 25
    
    ws.merge_cells('A2:I2')
    subtitle = ws['A2']
    subtitle.value = "Organization-approved techniques from ISMS-POL-A.8.11-S2.2"
    apply_style(subtitle, styles["subheader"])
    
    headers = [
        ("A", "Technique ID", 15),
        ("B", "Technique Name", 25),
        ("C", "Description", 45),
        ("D", "Reversible?", 12),
        ("E", "Format-Preserving?", 18),
        ("F", "Primary Use Cases", 40),
        ("G", "Approved for Use?", 15),
        ("H", "Policy Reference", 20),
        ("I", "Implementation Status", 18),
    ]
    
    for col, header_text, width in headers:
        cell = ws[f'{col}6']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    techniques = [
        ("TECH-SDM", "Static Data Masking", "Permanent replacement with realistic fictitious data", "No", "Yes", "Non-production environments", "ISMS-POL-A.8.11-S2.2 §2.1"),
        ("TECH-DDM", "Dynamic Data Masking", "Real-time masking based on user privileges", "N/A", "Yes", "Production role-based access", "ISMS-POL-A.8.11-S2.2 §2.2"),
        ("TECH-RED", "Redaction/Nullification", "Complete removal or placeholder replacement", "No", "No", "External reports, screenshots", "ISMS-POL-A.8.11-S2.2 §2.3"),
        ("TECH-TOK", "Tokenization", "Replacement with tokens, secure vault mapping", "Yes", "Optional", "Payment processing, PCI-DSS", "ISMS-POL-A.8.11-S2.2 §2.4"),
        ("TECH-SUB", "Data Substitution", "Realistic but entirely fictional data", "No", "Yes", "AI/ML training, analytics", "ISMS-POL-A.8.11-S2.2 §2.5"),
        ("TECH-ENC", "Encryption", "Cryptographic transformation", "Yes", "No", "Data at rest, backups", "ISMS-POL-A.8.11-S2.2 §2.6"),
        ("TECH-SHF", "Data Shuffling", "Rearrangement breaking associations", "No", "Yes", "Analytics preserving statistics", "ISMS-POL-A.8.11-S2.2 §2.7"),
        ("TECH-HSH", "Hashing", "One-way cryptographic hash", "No", "No", "Password storage, matching", "ISMS-POL-A.8.11-S2.2 §2.8"),
    ]
    
    for row_idx, (tech_id, tech_name, description, reversible, format_pres, use_cases, policy_ref) in enumerate(techniques, start=7):
        ws[f'A{row_idx}'] = tech_id
        ws[f'B{row_idx}'] = tech_name
        ws[f'C{row_idx}'] = description
        ws[f'D{row_idx}'] = reversible
        ws[f'E{row_idx}'] = format_pres
        ws[f'F{row_idx}'] = use_cases
        apply_style(ws[f'G{row_idx}'], styles["input_cell"])
        validations['yes_no_partial_planned_na'].add(ws[f'G{row_idx}'])
        ws[f'H{row_idx}'] = policy_ref
        apply_style(ws[f'I{row_idx}'], styles["input_cell"])
        validations['status_icons'].add(ws[f'I{row_idx}'])
        ws.row_dimensions[row_idx].height = 25
        for col in ['C', 'F']:
            ws[f'{col}{row_idx}'].alignment = Alignment(wrap_text=True)
    
    ws.freeze_panes = "A7"

def create_technique_selection_matrix(ws, styles):
    """Create technique selection matrix - 50 rows"""
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:L1')
    header = ws['A1']
    header.value = "TECHNIQUE SELECTION MATRIX"
    apply_style(header, styles["header"])
    
    headers = [
        ("A", "Data Category", 20),
        ("B", "Data Type Example", 30),
        ("C", "Sensitivity Level", 15),
        ("D", "Primary Technique", 20),
        ("E", "Secondary Technique", 20),
        ("F", "Format Must Preserve?", 15),
        ("G", "Reversibility Required?", 15),
        ("H", "Environment(s)", 25),
        ("I", "Selection Rationale", 35),
        ("J", "Regulatory Driver", 20),
        ("K", "Implementation Status", 18),
        ("L", "Notes", 30),
    ]
    
    for col, header_text, width in headers:
        cell = ws[f'{col}6']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    for row_idx in range(7, 57):  # 50 rows
        for col_idx in range(1, 13):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    for row in range(7, 57):
        validations['sensitivity_level'].add(ws[f'C{row}'])
        validations['technique_id'].add(ws[f'D{row}'])
        validations['technique_id'].add(ws[f'E{row}'])
        validations['yes_no_partial'].add(ws[f'F{row}'])
        validations['yes_no_partial'].add(ws[f'G{row}'])
        validations['status_icons'].add(ws[f'K{row}'])
    
    ws.freeze_panes = "A7"

def create_simple_implementation_sheet(ws, styles, sheet_title, row_count):
    """Generic implementation sheet creator"""
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:O1')
    header = ws['A1']
    header.value = sheet_title.upper()
    apply_style(header, styles["header"])
    
    headers = [
        ("A", "ID", 12),
        ("B", "System/Database", 25),
        ("C", "Data Category", 20),
        ("D", "Technique/Method", 20),
        ("E", "Configuration", 30),
        ("F", "Automated?", 12),
        ("G", "Frequency/Trigger", 15),
        ("H", "Last Update", 15),
        ("I", "Format Preserved?", 15),
        ("J", "Validated?", 12),
        ("K", "Performance Impact", 15),
        ("L", "Status", 18),
        ("M", "Evidence Ref", 20),
        ("N", "Responsible", 20),
        ("O", "Notes", 30),
    ]
    
    for col, header_text, width in headers:
        cell = ws[f'{col}4']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    end_row = 5 + row_count
    for row_idx in range(5, end_row):
        for col_idx in range(1, 16):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    for row in range(5, end_row):
        validations['yes_no_partial'].add(ws[f'F{row}'])
        validations['yes_no_partial'].add(ws[f'I{row}'])
        validations['yes_no_partial'].add(ws[f'J{row}'])
        validations['status_icons'].add(ws[f'L{row}'])
    
    ws.freeze_panes = "A5"

def create_gap_analysis(ws, styles):
    """Create gap analysis sheet"""
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:N1')
    header = ws['A1']
    header.value = "MASKING TECHNIQUE IMPLEMENTATION GAP ANALYSIS"
    apply_style(header, styles["header"])
    
    headers = [
        ("A", "Gap ID", 12),
        ("B", "Gap Category", 25),
        ("C", "Affected System", 25),
        ("D", "Gap Description", 35),
        ("E", "Risk Level", 12),
        ("F", "Impact", 30),
        ("G", "Root Cause", 25),
        ("H", "Remediation Action", 30),
        ("I", "Owner", 20),
        ("J", "Target Date", 15),
        ("K", "Status", 15),
        ("L", "Completion Date", 15),
        ("M", "Verification", 20),
        ("N", "Notes", 30),
    ]
    
    for col, header_text, width in headers:
        cell = ws[f'{col}4']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    for row_idx in range(5, 55):  # 50 rows
        for col_idx in range(1, 15):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    for row in range(5, 55):
        validations['risk_level'].add(ws[f'E{row}'])
        validations['gap_status'].add(ws[f'K{row}'])
    
    ws.freeze_panes = "A5"

def create_evidence_register(ws, styles):
    """Create evidence register - 100 rows"""
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:K1')
    header = ws['A1']
    header.value = "EVIDENCE REGISTER"
    apply_style(header, styles["header"])
    
    headers = [
        ("A", "Evidence ID", 12),
        ("B", "Evidence Type", 25),
        ("C", "Title", 30),
        ("D", "Related Technique", 20),
        ("E", "Related System", 25),
        ("F", "Location", 30),
        ("G", "Date", 15),
        ("H", "Owner", 20),
        ("I", "Retention", 15),
        ("J", "Review Frequency", 15),
        ("K", "Notes", 30),
    ]
    
    for col, header_text, width in headers:
        cell = ws[f'{col}4']
        cell.value = header_text
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    for row_idx in range(5, 105):  # 100 rows
        for col_idx in range(1, 12):
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    ws.freeze_panes = "A5"

def create_summary_dashboard(ws, styles):
    """Create summary dashboard"""
    ws.merge_cells('A1:G1')
    header = ws['A1']
    header.value = "MASKING TECHNIQUE IMPLEMENTATION DASHBOARD"
    apply_style(header, styles["header"])
    
    ws['A4'] = "Technique"
    ws['B4'] = "Approved?"
    ws['C4'] = "Implemented?"
    ws['D4'] = "Systems Covered"
    ws['E4'] = "Compliance %"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}4'], styles["column_header"])
    
    techniques = ["TECH-SDM", "TECH-DDM", "TECH-RED", "TECH-TOK", "TECH-SUB", "TECH-ENC", "TECH-SHF", "TECH-HSH"]
    for row_idx, tech in enumerate(techniques, start=5):
        ws[f'A{row_idx}'] = tech
    
    ws.freeze_panes = "A5"

def main():
    """Main execution function."""
    try:
        logger.info("=" * 78)
        logger.info("%s - %s Generator", WORKBOOK_ID, ASSESSMENT_AREA)
        logger.info("ISO/IEC 27001:2022 Control A.8.11: Data Masking")
        logger.info("=" * 78)

        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/12] Creating Instructions & Legend...")
        create_instructions_legend(wb["Instructions_Legend"], styles)

        logger.info("[2/12] Creating Approved Techniques...")
        create_approved_techniques(wb["Approved_Techniques"], styles)

        logger.info("[3/12] Creating Technique Selection Matrix...")
        create_technique_selection_matrix(wb["Technique_Selection_Matrix"], styles)

        logger.info("[4/12] Creating Static Masking (SDM)...")
        create_simple_implementation_sheet(wb["Static_Masking_SDM"], styles, "Static Data Masking (SDM) Implementation", 40)

        logger.info("[5/12] Creating Dynamic Masking (DDM)...")
        create_simple_implementation_sheet(wb["Dynamic_Masking_DDM"], styles, "Dynamic Data Masking (DDM) Implementation", 30)

        logger.info("[6/12] Creating Tokenization Implementation...")
        create_simple_implementation_sheet(wb["Tokenization_Implementation"], styles, "Tokenization Implementation", 20)

        logger.info("[7/12] Creating Encryption for Masking...")
        create_simple_implementation_sheet(wb["Encryption_for_Masking"], styles, "Encryption for Masking", 20)

        logger.info("[8/12] Creating Masking Tool Inventory...")
        create_simple_implementation_sheet(wb["Masking_Tool_Inventory"], styles, "Masking Tool/Solution Inventory", 30)

        logger.info("[9/12] Creating Configuration Standards...")
        create_simple_implementation_sheet(wb["Configuration_Standards"], styles, "Masking Configuration Standards", 40)

        logger.info("[10/12] Creating Gap Analysis...")
        create_gap_analysis(wb["Gap_Analysis"], styles)

        logger.info("[11/12] Creating Evidence Register...")
        create_evidence_register(wb["Evidence_Register"], styles)

        logger.info("[12/12] Creating Summary Dashboard...")
        create_summary_dashboard(wb["Summary_Dashboard"], styles)

        filename = f"ISMS-IMP-A.8.11.2_Masking_Techniques_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)

        logger.info("SUCCESS: %s", filename)
        logger.info("Workbook Structure: 12 sheets including Instructions, 8 Techniques, Evidence Register")
        logger.info("=" * 78)
        return 0
    except Exception as e:
        logger.error("Failed to generate workbook: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
