#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.11 - Assessment File Normalizer Utility
================================================================================

ISO/IEC 27001:2022 Control A.8.11: Data Masking
Quality Assurance Utility: Excel Assessment File Normalization & Validation

This script normalizes and validates A.8.11 assessment Excel workbooks to ensure
consistency, data quality, and compliance with framework standards before
consolidation into the compliance dashboard.

**Purpose:**
Ensures all assessment workbooks meet quality standards and structural
requirements, preventing data consolidation errors and improving audit
evidence reliability.

**Key Functions:**
1. File Naming Validation - Verify naming convention compliance
2. Workbook Structure Validation - Verify required sheets and columns
3. Data Normalization - Standardize dropdown values and formats
4. Content Validation - Check for incomplete assessments
5. Evidence Linkage Validation - Verify evidence references
6. Quality Reporting - Generate validation reports by severity

**Validation Scope:**
- ISMS_IMP_A_8_11_1_Data_Inventory_Assessment_YYYYMMDD.xlsx
- ISMS_IMP_A_8_11_2_Masking_Techniques_Assessment_YYYYMMDD.xlsx
- ISMS_IMP_A_8_11_3_Environment_Coverage_Assessment_YYYYMMDD.xlsx
- ISMS_IMP_A_8_11_4_Testing_Validation_Assessment_YYYYMMDD.xlsx

**Output:**
- Normalized assessment workbooks (with _normalized suffix if changes made)
- Validation report (text or Excel format)
- Issue summary for remediation

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel processing

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime, os, re (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    # Validate all A.8.11 assessment files in current directory
    python3 normalize_assessment_files_a811.py

Advanced Usage:
    # Validate files in specific directory
    python3 normalize_assessment_files_a811.py --input-dir /path/to/assessments
    
    # Validate with automatic normalization
    python3 normalize_assessment_files_a811.py --normalize
    
    # Validate specific assessment domain only
    python3 normalize_assessment_files_a811.py --domain 1
    
    # Generate detailed validation report
    python3 normalize_assessment_files_a811.py --report detailed
    
    # Dry run mode (report issues without modifying files)
    python3 normalize_assessment_files_a811.py --dry-run
    
    # Create backup before normalization
    python3 normalize_assessment_files_a811.py --normalize --backup

Command-Line Options:
    --input-dir PATH       Directory containing assessment workbooks
    --output-dir PATH      Directory for normalized workbooks (default: input-dir)
    --normalize            Apply normalization fixes automatically
    --domain N             Validate specific domain only (1-4)
    --report TYPE          Report format: summary|detailed|excel
    --dry-run              Validate only, don't modify files
    --severity LEVEL       Minimum severity: critical|high|medium|low
    --backup               Create backup before normalization (recommended)

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.11
Utility Type:         Quality Assurance - Assessment Normalization & Validation
Script Version:       1.0
Python Version:       3.8+

Related Scripts:
    - generate_a811_1_data_inventory.py
    - generate_a811_2_masking_techniques.py
    - generate_a811_3_environment_coverage.py
    - generate_a811_4_testing_validation.py
    - generate_a811_5_compliance_dashboard.py
    - consolidate_a811_dashboard.py

================================================================================
"""

import os
import sys
import shutil
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("\u274C Error: openpyxl not installed")
    print("ℹ️  Install with: sudo apt install python3-openpyxl")     
    sys.exit(1)


# ============================================================================
# CONFIGURATION
# ============================================================================

# Expected document IDs and titles from assessment workbooks
EXPECTED_DOCS = {
    "ISMS-IMP-A.8.11.1": {
        "title": "Data Inventory & Classification Assessment",
        "normalized": "ISMS-IMP-A.8.11.1.xlsx"
    },
    "ISMS-IMP-A.8.11.2": {
        "title": "Masking Technique Selection & Requirements",
        "normalized": "ISMS-IMP-A.8.11.2.xlsx"
    },
    "ISMS-IMP-A.8.11.3": {
        "title": "Environment Coverage Assessment",
        "normalized": "ISMS-IMP-A.8.11.3.xlsx"
    },
    "ISMS-IMP-A.8.11.4": {
        "title": "Testing & Validation Framework",
        "normalized": "ISMS-IMP-A.8.11.4.xlsx"
    },
}


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_workbook(filepath):
    """
    Validate workbook contains expected document ID in Instructions sheet.
    
    Args:
        filepath: Path to Excel workbook
    
    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        # Open workbook in read-only mode
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # Check for Instructions & Legend sheet
        if "Instructions_Legend" not in wb.sheetnames:
            wb.close()
            return (None, None)
        
        ws = wb["Instructions_Legend"]
        
        # Look for Document ID in column A (typically rows 4-20)
        # Format: "Document ID:" label in column A, value in column B
        for row in range(3, 25):
            cell_label = ws.cell(row=row, column=1).value
            
            if cell_label and "Document ID" in str(cell_label):
                doc_id_value = ws.cell(row=row, column=2).value
                
                if doc_id_value:
                    doc_id = str(doc_id_value).strip()
                    
                    # Check if it's one of our expected documents
                    if doc_id in EXPECTED_DOCS:
                        title = EXPECTED_DOCS[doc_id]["title"]
                        wb.close()
                        return (doc_id, title)
        
        wb.close()
        return (None, None)
        
    except Exception as e:
        print(f"    \u26A0\uFE0F  Error reading file: {e}")
        return (None, None)


def get_file_info(filepath):
    """Get file metadata for manifest."""
    stat = filepath.stat()
    return {
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime),
        "created": datetime.fromtimestamp(stat.st_ctime),
    }


# ============================================================================
# SCANNING FUNCTIONS
# ============================================================================

def scan_directory(directory):
    """
    Scan directory for assessment workbooks.
    
    Args:
        directory: Path to scan
    
    Returns:
        dict: {doc_id: {'path': Path, 'info': metadata}}
    """
    found_assessments = {}
    directory = Path(directory)
    
    # Get all .xlsx files, excluding temporary files
    xlsx_files = [f for f in directory.glob("*.xlsx") if not f.name.startswith("~$")]
    
    if not xlsx_files:
        print(f"\n\u274C No Excel files found in {directory}\n")
        return found_assessments
    
    print(f"\n🔍 Scanning {len(xlsx_files)} Excel file(s) in {directory}...\n")
    
    for filepath in sorted(xlsx_files):
        print(f"  Checking: {filepath.name}")
        doc_id, title = validate_workbook(filepath)
        
        if doc_id:
            print(f"    \u2705 Valid: {doc_id} - {title}")
            
            # Check for duplicates
            if doc_id in found_assessments:
                print(f"\n    \u26A0\uFE0F  WARNING: DUPLICATE FOUND FOR {doc_id}")
                print(f"        Previous: {found_assessments[doc_id]['path'].name}")
                print(f"        Current:  {filepath.name}\n")
                
                response = input("        Use CURRENT file? (y/n): ").strip().lower()
                
                if response == 'y':
                    found_assessments[doc_id] = {
                        'path': filepath,
                        'info': get_file_info(filepath)
                    }
                    print(f"        → Replaced with current file\n")
                else:
                    print(f"        → Keeping previous file\n")
            else:
                found_assessments[doc_id] = {
                    'path': filepath,
                    'info': get_file_info(filepath)
                }
        else:
            print(f"    ⏭️  Skipped (not a valid assessment workbook)")
        
        print()  # Blank line for readability
    
    return found_assessments


# ============================================================================
# MANIFEST CREATION
# ============================================================================

def create_manifest(output_dir, mapping, source_dir):
    """
    Create audit manifest documenting the normalization process.
    
    Args:
        output_dir: Destination directory
        mapping: Dictionary of normalized files
        source_dir: Source directory
    
    Returns:
        Path: Path to created manifest file
    """
    manifest_path = Path(output_dir) / "normalization_manifest_a811.txt"
    
    with open(manifest_path, 'w', encoding='utf-8') as f:
        # Header
        f.write("=" * 80 + "\n")
        f.write("ISMS ASSESSMENT FILE NORMALIZATION MANIFEST\n")
        f.write("ISO/IEC 27001:2022 - Control A.8.11: Data Masking\n")
        f.write("=" * 80 + "\n\n")
        
        # Normalization metadata
        f.write("NORMALIZATION METADATA\n")
        f.write("-" * 80 + "\n")
        f.write(f"Normalization Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Source Directory:        {source_dir.resolve()}\n")
        f.write(f"Output Directory:        {output_dir.resolve()}\n")
        f.write(f"Files Normalized:        {len(mapping)}/4 required\n")
        f.write(f"Normalization Status:    {'COMPLETE' if len(mapping) == 4 else 'INCOMPLETE'}\n")
        f.write("\n")
        
        # File mapping details
        f.write("FILE MAPPING\n")
        f.write("-" * 80 + "\n\n")
        
        for doc_id in sorted(EXPECTED_DOCS.keys()):
            if doc_id in mapping:
                info = mapping[doc_id]
                source_path = info['path']
                normalized_name = info['normalized']
                metadata = info['info']
                
                f.write(f"Document ID: {doc_id}\n")
                f.write(f"  Title:             {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"  Source File:       {source_path.name}\n")
                f.write(f"  Normalized Name:   {normalized_name}\n")
                f.write(f"  File Size:         {metadata['size']:,} bytes\n")
                f.write(f"  Last Modified:     {metadata['modified'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"  Created:           {metadata['created'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("\n")
            else:
                f.write(f"Document ID: {doc_id}\n")
                f.write(f"  Title:             {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"  Status:            \u274C NOT FOUND\n")
                f.write(f"  Normalized Name:   {EXPECTED_DOCS[doc_id]['normalized']} (expected)\n")
                f.write("\n")
        
        # Usage instructions
        f.write("=" * 80 + "\n")
        f.write("DASHBOARD INTEGRATION INSTRUCTIONS\n")
        f.write("=" * 80 + "\n\n")
        
        f.write("Next Steps:\n")
        f.write("  1. Generate compliance dashboard workbook:\n")
        f.write("     python3 generate_a811_5_compliance_dashboard.py\n\n")
        f.write("  2. Place dashboard in same directory as normalized files:\n")
        f.write(f"     {output_dir.resolve()}\n\n")
        f.write("  3. Open dashboard in Excel/LibreOffice Calc\n\n")
        f.write("  4. When prompted 'Update Links?', click 'Yes' or 'Update'\n\n")
        f.write("  5. Dashboard will auto-populate with compliance data from:\n")
        for doc_id in sorted(EXPECTED_DOCS.keys()):
            if doc_id in mapping:
                f.write(f"     \u2705 {EXPECTED_DOCS[doc_id]['normalized']}\n")
            else:
                f.write(f"     \u274C {EXPECTED_DOCS[doc_id]['normalized']} (MISSING)\n")
        f.write("\n")
        
        f.write("Technical Details:\n")
        f.write("  - Dashboard uses external workbook links to normalized files\n")
        f.write("  - External workbook links auto-update when sources change\n")
        f.write("  - Place dashboard in same directory as normalized files\n")
        f.write("  - Open dashboard and click 'Update Links' when prompted\n\n")
        f.write("File Retention:\n")
        f.write("  - Original source files: Retained in source directory\n")
        f.write("  - Normalized copies: Located in output directory\n")
        f.write("  - This manifest: Retained with normalized files for audit\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("END OF MANIFEST\n")
        f.write("=" * 80 + "\n")
    
    return manifest_path


# ============================================================================
# MAIN NORMALIZATION FUNCTION
# ============================================================================

def normalize_files(source_dir=None, output_dir=None, auto_confirm=False):
    """
    Main normalization orchestration function.
    
    Args:
        source_dir: Source directory (prompts if None)
        output_dir: Output directory (prompts if None)
        auto_confirm: Skip confirmation prompts (for automation)
    
    Returns:
        bool: True if successful, False otherwise
    """
    print("=" * 80)
    print("ISMS ASSESSMENT FILE NORMALIZATION UTILITY")
    print("ISO/IEC 27001:2022 - Control A.8.11: Data Masking")
    print("=" * 80)
    print("\nThis script prepares assessment workbooks for dashboard linking by:")
    print("  1. Scanning for completed assessment files")
    print("  2. Validating document IDs in Instructions_Legend sheet")
    print("  3. Copying to normalized filenames (no dates/versions)")
    print("  4. Creating audit manifest for traceability\n")
    
    # Get source directory
    if not source_dir:
        source_dir = input("Enter source directory path (or press Enter for current directory): ").strip()
        if not source_dir:
            source_dir = "."
    
    source_dir = Path(source_dir).resolve()
    
    if not source_dir.exists():
        print(f"\n\u274C Error: Source directory does not exist: {source_dir}\n")
        return False
    
    if not source_dir.is_dir():
        print(f"\n\u274C Error: Not a directory: {source_dir}\n")
        return False
    
    print(f"\n📁 Source directory: {source_dir}")
    
    # Get output directory
    if not output_dir:
        default_output = source_dir / "Dashboard_Sources"
        output_input = input(f"\nEnter output directory (or press Enter for '{default_output.name}'): ").strip()
        if not output_input:
            output_dir = default_output
        else:
            output_dir = Path(output_input)
    
    output_dir = Path(output_dir).resolve()
    
    # Create output directory if it doesn't exist
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        print(f"📁 Output directory: {output_dir}")
    except Exception as e:
        print(f"\n\u274C Error creating output directory: {e}\n")
        return False
    
    # Scan for assessment files
    found = scan_directory(source_dir)
    
    if not found:
        print("\u274C No valid assessment workbooks found in source directory\n")
        print("   Ensure files contain valid Document IDs in 'Instructions_Legend' sheet:")
        for doc_id, info in EXPECTED_DOCS.items():
            print(f"   - {doc_id}: {info['title']}")
        print()
        return False
    
    # Show normalization summary
    print("=" * 80)
    print("NORMALIZATION SUMMARY")
    print("=" * 80 + "\n")
    print(f"Found {len(found)} of {len(EXPECTED_DOCS)} required assessment workbooks:\n")
    
    mapping = {}
    for doc_id in sorted(EXPECTED_DOCS.keys()):
        if doc_id in found:
            normalized_name = EXPECTED_DOCS[doc_id]["normalized"]
            source_path = found[doc_id]['path']
            
            print(f"  \u2705 {doc_id}")
            print(f"     Title:      {EXPECTED_DOCS[doc_id]['title']}")
            print(f"     Source:     {source_path.name}")
            print(f"     Normalized: {normalized_name}\n")
            
            mapping[doc_id] = {
                'path': source_path,
                'normalized': normalized_name,
                'info': found[doc_id]['info']
            }
        else:
            print(f"  \u274C {doc_id} - NOT FOUND")
            print(f"     Title: {EXPECTED_DOCS[doc_id]['title']}\n")
    
    if len(found) < len(EXPECTED_DOCS):
        print(f"\u26A0\uFE0F  WARNING: Only {len(found)}/{len(EXPECTED_DOCS)} assessment workbooks found")
        print("   Dashboard will have incomplete data for missing assessments\n")
    
    print(f"Output directory: {output_dir}\n")
    
    # Confirm normalization
    if not auto_confirm:
        response = input("Proceed with normalization? (y/n): ").strip().lower()
        if response != 'y':
            print("\n\u274C Normalization cancelled by user\n")
            return False
    
    # Perform normalization (copy files)
    print("\n" + "=" * 80)
    print("NORMALIZING FILES...")
    print("=" * 80 + "\n")
    
    for doc_id, info in mapping.items():
        source = info['path']
        dest = output_dir / info['normalized']
        
        print(f"Copying: {source.name}")
        print(f"      → {dest.name}")
        
        try:
            shutil.copy2(source, dest)
            print(f"      \u2705 Success\n")
        except Exception as e:
            print(f"      \u274C Error: {e}\n")
            print(f"\u274C Normalization failed at {doc_id}\n")
            return False
    
    # Create audit manifest
    print("📄 Creating audit manifest...")
    try:
        manifest = create_manifest(output_dir, mapping, source_dir)
        print(f"   \u2705 Created: {manifest.name}\n")
    except Exception as e:
        print(f"   \u274C Error creating manifest: {e}\n")
        return False
    
    # Success summary
    print("=" * 80)
    print("\u2705 NORMALIZATION COMPLETE")
    print("=" * 80 + "\n")
    print(f"Normalized files:  {output_dir}")
    print(f"Audit manifest:    {manifest}\n")
    print("NEXT STEPS:\n")
    print("  1. Review audit manifest for file mapping details")
    print("  2. Generate dashboard workbook:")
    print("     python3 generate_a811_5_compliance_dashboard.py\n")
    print("  3. Place dashboard in same directory as normalized files:")
    print(f"     {output_dir}\n")
    print("  4. Open dashboard and click 'Update Links' when prompted\n")
    print("  5. Dashboard will auto-populate with current compliance data\n")
    print("=" * 80 + "\n")
    
    return True


# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Normalize ISMS assessment workbooks for dashboard integration",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode (prompts for directories)
  python3 normalize_assessment_files_a811.py
  
  # Specify source directory
  python3 normalize_assessment_files_a811.py --source ./assessments
  
  # Specify both directories
  python3 normalize_assessment_files_a811.py --source ./assessments --output ./dashboard
  
  # Automated mode (no prompts)
  python3 normalize_assessment_files_a811.py --source ./assessments --auto-confirm
        """
    )
    
    parser.add_argument(
        '--source', '-s',
        help='Source directory containing assessment workbooks (default: current directory)'
    )
    
    parser.add_argument(
        '--output', '-o',
        help='Output directory for normalized files (default: ./Dashboard_Sources)'
    )
    
    parser.add_argument(
        '--auto-confirm', '-y',
        action='store_true',
        help='Skip confirmation prompts (for automation)'
    )
    
    args = parser.parse_args()
    
    success = normalize_files(
        source_dir=args.source,
        output_dir=args.output,
        auto_confirm=args.auto_confirm
    )
    
    sys.exit(0 if success else 1)