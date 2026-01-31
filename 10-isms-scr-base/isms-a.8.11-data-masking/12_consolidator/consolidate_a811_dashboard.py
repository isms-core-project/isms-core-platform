#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.11.5 - Dashboard Data Consolidation Script
================================================================================

ISO/IEC 27001:2022 Control A.8.11: Data Masking
Alternative Data Integration Method for Compliance Dashboard

This utility provides an alternative method for populating the A.8.11.5
Compliance Dashboard with data from source assessment workbooks by directly
copying user-entered data rather than using external workbook formula references.

**Primary Approach vs. Alternative Approach:**

**Primary (Recommended): External Workbook Formula Linking**
- Dashboard contains formulas referencing source workbooks
- Data updates automatically when source files change
- Dashboard and source files must be in same directory
- Requires "Update Links" when opening dashboard
- See: generate_a811_5_compliance_dashboard.py

**Alternative (This Script): Direct Data Consolidation**
- Reads data from source workbooks programmatically
- Copies data values (not formulas) into dashboard
- Creates static snapshot of compliance status
- Works with dashboard and sources in different locations
- Requires manual re-run when source data changes

**When to Use This Script:**
- Excel security settings prevent external workbook links
- SharePoint/OneDrive environments where linking fails
- Email distribution requires self-contained workbooks
- Monthly/quarterly compliance reporting archives
- Board presentations requiring frozen data
- Audit evidence packages with point-in-time data

**Supported Source Workbooks:**
- ISMS-IMP-A.8.11.1.xlsx (Data Inventory & Classification)
- ISMS-IMP-A.8.11.2.xlsx (Masking Technique Selection)
- ISMS-IMP-A.8.11.3.xlsx (Environment Coverage)
- ISMS-IMP-A.8.11.4.xlsx (Testing & Validation)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel manipulation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Excel file reading and writing)
    - sys, os, datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 consolidate_a811_dashboard.py ISMS-IMP-A.8.11.5_Dashboard_YYYYMMDD.xlsx

Prerequisites:
    1. Generate dashboard first:
       python3 generate_a811_5_compliance_dashboard.py
    
    2. Normalize source assessment files:
       python3 normalize_assessment_files_a811.py
    
    3. Ensure all files are accessible:
       - ISMS-IMP-A.8.11.1.xlsx
       - ISMS-IMP-A.8.11.2.xlsx
       - ISMS-IMP-A.8.11.3.xlsx
       - ISMS-IMP-A.8.11.4.xlsx
       - Dashboard file (provided as argument)

File Location Requirements:
    Source workbooks must be in CURRENT DIRECTORY when running script.
    Dashboard can be in any directory (specify full path if needed).

Exit Codes:
    0 = Consolidation successful
    1 = Consolidation failed (missing files, errors during processing)

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.11
Script Type:          Data Consolidation / Alternative Integration Method
Script Version:       1.0
Python Version:       3.8+

Related Scripts:
    - generate_a811_5_compliance_dashboard.py (creates dashboard)
    - normalize_assessment_files_a811.py (normalizes file names)
    - excel_sanity_check_a811.py (workbook validation)

================================================================================
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from datetime import datetime, timedelta

SOURCE_WORKBOOKS = {
    "wb1": "ISMS-IMP-A.8.11.1.xlsx",
    "wb2": "ISMS-IMP-A.8.11.2.xlsx",
    "wb3": "ISMS-IMP-A.8.11.3.xlsx",
    "wb4": "ISMS-IMP-A.8.11.4.xlsx",
}

DOMAIN_NAMES = {
    "wb1": "Data Inventory",
    "wb2": "Masking Techniques",
    "wb3": "Environment Coverage",
    "wb4": "Testing Validation",
}

SHEET_START_ROWS = {
    "Consolidated_Gap_Analysis": 6,
    "Risk_Register": 9,
    "Remediation_Roadmap": 4,
    "Evidence_Master_Index": 4,
}

def safely_write_data(ws, start_row, data):
    """Safely write data, handling merged cells"""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except:
                continue
        entries_written += 1
    return entries_written

def extract_gaps_from_workbook(filepath, domain_name):
    """Extract gap items from source workbook"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        gaps = []
        gap_counter = 1
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['Instructions', 'Summary_Dashboard', 'Evidence_Register', 
                             'Approval_Sign_Off', 'Document_Control']:
                continue
            
            ws = wb[sheet_name]
            
            for row in range(8, min(ws.max_row + 1, 100)):
                row_values = [ws.cell(row=row, column=col).value for col in range(1, 20)]
                
                status = None
                status_col = None
                for col_idx, val in enumerate(row_values, start=1):
                    if val in ['[!] Partial', '[X] Non-Compliant']:
                        status = val
                        status_col = col_idx
                        break
                
                if not status:
                    continue
                
                system_name = row_values[0] if row_values[0] else 'Unknown'
                gap_desc = None
                
                if status_col:
                    for offset in [2, 3, 4]:
                        potential_gap = ws.cell(row=row, column=status_col + offset).value
                        if potential_gap and len(str(potential_gap)) > 10:
                            gap_desc = potential_gap
                            break
                
                severity = 'High' if status == '[X] Non-Compliant' else 'Medium'
                
                gap_entry = [
                    domain_name,
                    f'GAP-{domain_name.split()[0]}-{gap_counter:03d}',
                    gap_desc if gap_desc else f'{system_name}: Configuration gap',
                    severity,
                    status.replace('[!] ', '').replace('[X] ', ''),
                    'Compliant',
                    None,
                    'Open',
                ]
                gaps.append(gap_entry)
                gap_counter += 1
        
        wb.close()
        return gaps
    except Exception as e:
        print(f"  [!] Error extracting gaps from {filepath}: {e}")
        return []

def extract_evidence_from_workbook(filepath, domain_name):
    """Extract evidence entries from source workbook"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        evidence_sheet_names = ['Evidence_Register', 'Evidence_Master_Index']
        ws_name = None
        for name in evidence_sheet_names:
            if name in wb.sheetnames:
                ws_name = name
                break
        
        if not ws_name:
            wb.close()
            return []
        
        ws = wb[ws_name]
        evidence = []
        
        for row in range(10, ws.max_row + 1):
            evidence_id = ws[f'A{row}'].value
            if evidence_id and str(evidence_id).startswith('EVD-'):
                row_data = [
                    evidence_id,
                    domain_name,
                    ws[f'B{row}'].value,
                    ws[f'C{row}'].value,
                    ws[f'D{row}'].value,
                    ws[f'E{row}'].value,
                    ws[f'F{row}'].value,
                    ws[f'G{row}'].value,
                ]
                evidence.append(row_data)
        
        wb.close()
        return evidence
    except Exception as e:
        print(f"  [!] Error reading evidence from {filepath}: {e}")
        return []

def generate_risks_from_gaps(gaps):
    """Generate risk entries from critical gaps"""
    risks = []
    risk_counter = 1
    
    for gap in gaps:
        domain = gap[0]
        gap_id = gap[1]
        gap_desc = gap[2]
        severity = gap[3]
        
        if severity != 'High':
            continue
        
        risk_entry = [
            f'RISK-{risk_counter:03d}',
            gap_id,
            'Security',
            f'Security risk: {gap_desc[:80]}...',
            domain,
            'High',
            'High',
            'High',
            'Open',
            f'Mitigate gap {gap_id}',
            None,
            None,
            None,
        ]
        risks.append(risk_entry)
        risk_counter += 1
    
    return risks

def generate_remediation_from_gaps(gaps):
    """Generate remediation roadmap from gaps"""
    remediation = []
    action_counter = 1
    today = datetime.now()
    
    for gap in gaps:
        domain = gap[0]
        gap_id = gap[1]
        gap_desc = gap[2]
        severity = gap[3]
        
        priority = 'Critical' if severity == 'High' else 'High' if severity == 'Medium' else 'Medium'
        days_offset = 30 if priority == 'Critical' else 60 if priority == 'High' else 90
        target_date = (today + timedelta(days=days_offset)).strftime('%Y-%m-%d')
        
        remediation_entry = [
            f'REM-{action_counter:03d}',
            gap_id,
            domain,
            f'Remediate: {gap_desc[:60]}...',
            priority,
            'Planned',
            None,
            today.strftime('%Y-%m-%d'),
            target_date,
            None,
            None,
            None,
        ]
        remediation.append(remediation_entry)
        action_counter += 1
    
    return remediation

def populate_dashboard(dashboard_file):
    """Populate dashboard from source workbooks"""
    
    print("="*80)
    print("ISMS A.8.11 DATA MASKING DASHBOARD CONSOLIDATION")
    print("="*80)
    
    missing = []
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        if not os.path.exists(wb_filename):
            missing.append(wb_filename)
    
    if missing:
        print("\n[X] ERROR: Missing source workbooks:")
        for wb in missing:
            print(f"    - {wb}")
        print("\n  Place all 4 assessment workbooks in same directory.")
        return False
    
    all_gaps = []
    all_evidence = []
    
    print("\n[1/4] Extracting gaps from source workbooks...")
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        domain_name = DOMAIN_NAMES[wb_key]
        print(f"  * {domain_name}: ", end="")
        gaps = extract_gaps_from_workbook(wb_filename, domain_name)
        print(f"{len(gaps)} gaps")
        all_gaps.extend(gaps)
    
    print(f"\n  Total gaps identified: {len(all_gaps)}")
    
    print("\n[2/4] Extracting evidence from source workbooks...")
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        domain_name = DOMAIN_NAMES[wb_key]
        print(f"  * {domain_name}: ", end="")
        evidence = extract_evidence_from_workbook(wb_filename, domain_name)
        print(f"{len(evidence)} evidence docs")
        all_evidence.extend(evidence)
    
    print(f"\n  Total evidence collected: {len(all_evidence)}")
    
    print("\n[3/4] Generating risks and remediation...")
    all_risks = generate_risks_from_gaps(all_gaps)
    all_remediation = generate_remediation_from_gaps(all_gaps)
    print(f"  * Risks generated: {len(all_risks)}")
    print(f"  * Remediation actions: {len(all_remediation)}")
    
    print(f"\n[4/4] Writing to dashboard...")
    try:
        wb = load_workbook(dashboard_file)
    except Exception as e:
        print(f"  [X] Error loading dashboard: {e}")
        return False
    
    if 'Consolidated_Gap_Analysis' in wb.sheetnames and all_gaps:
        ws = wb['Consolidated_Gap_Analysis']
        count = safely_write_data(ws, 6, all_gaps)
        print(f"  [OK] Consolidated_Gap_Analysis: {count} entries")
    
    if 'Risk_Register' in wb.sheetnames and all_gaps:
        ws = wb['Risk_Register']
        count = safely_write_data(ws, 9, all_gaps)
        print(f"  [OK] Risk_Register: {count} entries")
    
    if 'Remediation_Roadmap' in wb.sheetnames and all_gaps:
        ws = wb['Remediation_Roadmap']
        count = safely_write_data(ws, 4, all_gaps)
        print(f"  [OK] Remediation_Roadmap: {count} entries")
    
    if 'Evidence_Master_Index' in wb.sheetnames and all_gaps:
        ws = wb['Evidence_Master_Index']
        count = safely_write_data(ws, 4, all_gaps)
        print(f"  [OK] Evidence_Master_Index: {count} entries")
    
    try:
        wb.save(dashboard_file)
        print(f"\n[SAVED] {dashboard_file}")
    except Exception as e:
        print(f"  [X] Error saving: {e}")
        return False
    
    print("\n" + "="*80)
    print("[OK] DASHBOARD CONSOLIDATION COMPLETE")
    print("="*80)
    print(f"\nSummary:")
    print(f"  * Gaps: {len(all_gaps)}")
    print(f"  * Risks: {len(all_risks)}")
    print(f"  * Remediation: {len(all_remediation)}")
    print(f"  * Evidence: {len(all_evidence)}")
    print(f"\n[ROCKET] Dashboard ready!")
    print("="*80 + "\n")
    
    return True

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 consolidate_a811_dashboard.py <dashboard.xlsx>")
        sys.exit(1)
    
    dashboard_file = sys.argv[1]
    
    if not os.path.exists(dashboard_file):
        print(f"[X] Error: Dashboard file not found: {dashboard_file}")
        sys.exit(1)
    
    success = populate_dashboard(dashboard_file)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
