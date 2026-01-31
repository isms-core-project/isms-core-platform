#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.7-8 Workbook Sanity Check Script
================================================================================

Validates A.6.7-8 assessment workbooks for:
- Required sheets present
- Proper header formatting
- Data validation rules applied
- No corrupt or missing data
- Metadata consistency

Usage:
    python3 sanity_check_a678.py <workbook_path>
    python3 sanity_check_a678.py --all  # Check all workbooks in 90_workbooks

================================================================================
"""

import logging
import sys
import os
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# EXPECTED WORKBOOK STRUCTURES
# =============================================================================
EXPECTED_SHEETS = {
    "S1": ["Instructions", "Eligibility_Criteria", "Authorization_Register", "Risk_Assessment",
           "Physical_Security", "Acknowledgments", "Gap_Analysis", "Evidence_Register",
           "Dashboard", "Approval_Sign_Off"],
    "S2": ["Instructions", "VPN_Assessment", "MFA_Assessment", "Encryption_Assessment",
           "Logging_Assessment", "Compliance_Summary", "Gap_Analysis", "Evidence_Register",
           "Dashboard", "Approval_Sign_Off"],
    "S3": ["Instructions", "Device_Inventory", "Encryption_Status", "Endpoint_Protection",
           "Patch_Compliance", "BYOD_Assessment", "Physical_Security", "Lost_Stolen_Procedures",
           "Gap_Analysis", "Evidence_Register", "Dashboard", "Approval_Sign_Off"],
    "S4": ["Instructions", "Channel_Assessment", "Channel_Availability", "Event_Categories",
           "Response_Timeframes", "NonBlame_Culture", "Awareness_Training", "Sample_Events",
           "Gap_Analysis", "Evidence_Register", "Dashboard", "Approval_Sign_Off"],
    "S5": ["Instructions", "Executive_Summary", "Assessment_Status", "A67_Compliance",
           "A68_Compliance", "Gap_Consolidation", "Evidence_Summary", "Trend_Analysis",
           "Recommendations", "Approval_Sign_Off"],
}


class SanityCheckResult:
    """Container for sanity check results."""

    def __init__(self, workbook_path: str):
        self.workbook_path = workbook_path
        self.errors = []
        self.warnings = []
        self.info = []

    def add_error(self, message: str):
        self.errors.append(message)
        logger.error(f"  ERROR: {message}")

    def add_warning(self, message: str):
        self.warnings.append(message)
        logger.warning(f"  WARNING: {message}")

    def add_info(self, message: str):
        self.info.append(message)
        logger.info(f"  INFO: {message}")

    @property
    def passed(self) -> bool:
        return len(self.errors) == 0

    def summary(self) -> str:
        status = "PASSED" if self.passed else "FAILED"
        return f"{status}: {len(self.errors)} errors, {len(self.warnings)} warnings"


def detect_workbook_type(filename: str) -> str:
    """Detect which S1-S5 workbook type based on filename."""
    filename_lower = filename.lower()

    if ".s1" in filename_lower or "authorization" in filename_lower:
        return "S1"
    elif ".s2" in filename_lower or "technical" in filename_lower:
        return "S2"
    elif ".s3" in filename_lower or "endpoint" in filename_lower:
        return "S3"
    elif ".s4" in filename_lower or "event_reporting" in filename_lower or "reporting_mechanism" in filename_lower:
        return "S4"
    elif ".s5" in filename_lower or "dashboard" in filename_lower or "compliance_dashboard" in filename_lower:
        return "S5"
    else:
        return None


def check_workbook(filepath: str) -> SanityCheckResult:
    """
    Perform sanity checks on a workbook.

    Args:
        filepath: Path to the Excel workbook

    Returns:
        SanityCheckResult with check outcomes
    """
    result = SanityCheckResult(filepath)
    filename = os.path.basename(filepath)

    logger.info(f"Checking: {filename}")

    # Check file exists
    if not os.path.exists(filepath):
        result.add_error(f"File not found: {filepath}")
        return result

    # Detect workbook type
    wb_type = detect_workbook_type(filename)
    if not wb_type:
        result.add_warning("Could not detect workbook type (S1-S5) from filename")
    else:
        result.add_info(f"Detected workbook type: {wb_type}")

    try:
        wb = load_workbook(filepath, data_only=True)

        # Check sheet presence
        if wb_type and wb_type in EXPECTED_SHEETS:
            expected = EXPECTED_SHEETS[wb_type]
            actual = wb.sheetnames

            missing = set(expected) - set(actual)
            extra = set(actual) - set(expected)

            if missing:
                result.add_error(f"Missing sheets: {', '.join(missing)}")
            if extra:
                result.add_warning(f"Extra sheets found: {', '.join(extra)}")
            if not missing:
                result.add_info(f"All {len(expected)} expected sheets present")

        # Check Instructions sheet
        if "Instructions" in wb.sheetnames:
            ws = wb["Instructions"]

            # Check header cell
            header_cell = ws["A1"]
            if header_cell.value:
                if "ISMS-IMP-A.6.7-8" in str(header_cell.value):
                    result.add_info("Header contains correct document ID")
                else:
                    result.add_warning("Header may not contain correct document ID")
            else:
                result.add_error("Instructions header (A1) is empty")
        else:
            result.add_error("Instructions sheet not found")

        # Check Approval_Sign_Off sheet
        if "Approval_Sign_Off" in wb.sheetnames:
            ws = wb["Approval_Sign_Off"]
            if ws["A1"].value:
                result.add_info("Approval sheet has header")
            else:
                result.add_warning("Approval sheet header is empty")

        # Check for data validation (basic check)
        validation_found = False
        for ws in wb.worksheets:
            if ws.data_validations.dataValidation:
                validation_found = True
                break

        if validation_found:
            result.add_info("Data validations present in workbook")
        else:
            result.add_warning("No data validations found in workbook")

        wb.close()

    except Exception as e:
        result.add_error(f"Error loading workbook: {e}")

    return result


def check_all_workbooks(directory: str) -> list:
    """
    Check all A.6.7-8 workbooks in a directory.

    Args:
        directory: Path to directory containing workbooks

    Returns:
        List of SanityCheckResult objects
    """
    results = []

    workbooks_dir = Path(directory)
    if not workbooks_dir.exists():
        logger.error(f"Directory not found: {directory}")
        return results

    xlsx_files = list(workbooks_dir.glob("ISMS-IMP-A.6.7-8*.xlsx"))

    if not xlsx_files:
        logger.warning(f"No A.6.7-8 workbooks found in {directory}")
        return results

    logger.info(f"Found {len(xlsx_files)} workbooks to check")
    logger.info("=" * 60)

    for xlsx_file in xlsx_files:
        result = check_workbook(str(xlsx_file))
        results.append(result)
        logger.info(f"  Result: {result.summary()}")
        logger.info("-" * 40)

    return results


def main():
    """Main entry point."""
    logger.info("=" * 60)
    logger.info("ISMS-IMP-A.6.7-8 Workbook Sanity Check")
    logger.info("=" * 60)

    if len(sys.argv) < 2:
        print("Usage: python3 sanity_check_a678.py <workbook_path>")
        print("       python3 sanity_check_a678.py --all")
        sys.exit(1)

    if sys.argv[1] == "--all":
        # Check all workbooks in default directory
        script_dir = Path(__file__).parent
        workbooks_dir = script_dir.parent / "90_workbooks"

        results = check_all_workbooks(str(workbooks_dir))

        # Summary
        passed = sum(1 for r in results if r.passed)
        failed = len(results) - passed

        logger.info("=" * 60)
        logger.info(f"SANITY CHECK COMPLETE: {passed} passed, {failed} failed")
        logger.info("=" * 60)

        sys.exit(0 if failed == 0 else 1)
    else:
        # Check specific file
        filepath = sys.argv[1]
        result = check_workbook(filepath)

        logger.info("=" * 60)
        logger.info(f"Result: {result.summary()}")
        logger.info("=" * 60)

        sys.exit(0 if result.passed else 1)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation for A.6.7-8 workbook sanity checking
# =============================================================================
