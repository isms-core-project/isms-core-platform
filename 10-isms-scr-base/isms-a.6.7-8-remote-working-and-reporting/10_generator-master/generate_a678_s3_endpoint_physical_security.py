#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.7-8.S3 - Endpoint and Physical Security Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.6.7 (Remote Working) & A.6.8 (Event Reporting)
Assessment Domain S3 of S5: Endpoint and Physical Security Assessment

Reference Pattern: Based on ISMS-IMP-A.6.7-8.S3 specification
================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.7-8.S3"
WORKBOOK_NAME = "Endpoint and Physical Security Assessment"
CONTROL_ID = "A.6.7-8"
CONTROL_NAME = "Remote Working and Security Event Reporting"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header": {"font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
                   "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
                   "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)},
        "column_header": {"font": Font(name="Calibri", size=10, bold=True),
                          "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
                          "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
                          "border": border_thin},
        "input_cell": {"fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
                       "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
                       "border": border_thin},
        "border": border_thin,
    }

# =============================================================================
# PRE-POPULATED DATA
# =============================================================================
DEVICE_TYPES = ["Laptop", "Desktop", "Tablet", "Mobile", "Other"]
OWNERSHIP_TYPES = ["Corporate", "BYOD", "Contractor"]
ENCRYPTION_TYPES = ["BitLocker", "FileVault", "LUKS", "Other", "None"]
PHYSICAL_REQUIREMENTS = [
    ("Screen Privacy", "Position screen away from windows and common areas", "Mandatory"),
    ("Privacy Screen", "Use privacy screen filter in public spaces", "Conditional"),
    ("Secure Storage", "Lockable cabinet or drawer for sensitive documents", "Mandatory"),
    ("Device Security", "Cable lock or secure storage for portable devices", "Recommended"),
    ("Unattended Devices", "Never leave devices unattended in public", "Mandatory"),
    ("Document Handling", "Shred sensitive documents (cross-cut shredder)", "Mandatory"),
    ("Clear Desk", "Clear desk when leaving workspace", "Mandatory"),
    ("Access Control", "Prevent family/visitor access to work devices", "Mandatory"),
    ("Travel - Hand Luggage", "Carry devices in hand luggage", "Mandatory"),
    ("Travel - Vehicle", "Never leave devices in vehicles", "Mandatory"),
]
LOST_STOLEN_PROCEDURES = [
    ("Reporting Channel", "24/7 hotline or email for reporting", "1 hour"),
    ("Account Disable", "Ability to disable account within 1 hour", "1 hour"),
    ("Remote Lock", "Ability to lock device remotely within 1 hour", "1 hour"),
    ("Remote Wipe", "Ability to wipe device remotely within 4 hours", "4 hours"),
    ("Recovery Key Access", "Access to recovery keys for encrypted devices", "As needed"),
    ("Replacement Process", "Process to issue replacement device", "24-48 hours"),
    ("Incident Documentation", "Record of incident details", "Immediate"),
    ("Investigation", "Follow-up investigation procedure", "72 hours"),
]


def create_workbook() -> Workbook:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    sheets = ["Instructions", "Device_Inventory", "Encryption_Status", "Endpoint_Protection",
              "Patch_Compliance", "BYOD_Assessment", "Physical_Security", "Lost_Stolen_Procedures",
              "Gap_Analysis", "Evidence_Register", "Dashboard", "Approval_Sign_Off"]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def create_instructions_sheet(ws, styles):
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    doc_info = [("Document ID", DOCUMENT_ID),
                ("Assessment Area", "Device Security, Encryption, Endpoint Protection, Physical Security"),
                ("Related Policy", "ISMS-POL-A.6.7-8, Sections 2.2 (Physical) & 2.5 (Device)"),
                ("Version", "1.0"), ("Assessment Date", ""), ("Completed By", "")]
    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55


def create_device_inventory_sheet(ws, styles):
    ws.merge_cells("A1:K1")
    ws["A1"] = "Remote Device Inventory"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Device ID", "Device Type", "Ownership", "OS", "OS Version", "Assigned User",
               "Department", "Remote Enabled", "MDM Enrolled", "Last Check-in", "Status"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    for data_row in range(4, 24):
        ws.cell(row=data_row, column=1, value=f"DEV-{data_row-3:04d}").border = styles["border"]
        for col in range(2, 12):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    type_dv = DataValidation(type="list", formula1=f'"{",".join(DEVICE_TYPES)}"', allow_blank=True)
    ws.add_data_validation(type_dv)
    type_dv.add("B4:B23")

    own_dv = DataValidation(type="list", formula1=f'"{",".join(OWNERSHIP_TYPES)}"', allow_blank=True)
    ws.add_data_validation(own_dv)
    own_dv.add("C4:C23")

    yn_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yn_dv)
    yn_dv.add("H4:I23")

    status_dv = DataValidation(type="list", formula1='"Active,Inactive,Lost,Retired"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("K4:K23")

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 14


def create_encryption_status_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "Device Encryption Status"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    headers = ["Device ID", "Encryption Type", "Full-Disk", "Status", "Key Escrowed", "Last Verified", "Compliant", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    for data_row in range(4, 24):
        for col in range(1, 9):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    enc_dv = DataValidation(type="list", formula1=f'"{",".join(ENCRYPTION_TYPES)}"', allow_blank=True)
    ws.add_data_validation(enc_dv)
    enc_dv.add("B4:B23")

    yn_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yn_dv)
    yn_dv.add("C4:C23")
    yn_dv.add("E4:E23")
    yn_dv.add("G4:G23")

    status_dv = DataValidation(type="list", formula1='"Active,Suspended,Not Configured"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("D4:D23")

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15


def create_endpoint_protection_sheet(ws, styles):
    ws.merge_cells("A1:J1")
    ws["A1"] = "Endpoint Protection Status"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    headers = ["Device ID", "Protection Solution", "Agent Installed", "Agent Version", "Status",
               "Definitions Date", "Last Scan", "Threats (90d)", "Compliant", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    for data_row in range(4, 24):
        for col in range(1, 11):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14


def create_patch_compliance_sheet(ws, styles):
    ws.merge_cells("A1:I1")
    ws["A1"] = "Patch Compliance Status"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    headers = ["Device ID", "OS Patch Level", "Critical Missing", "High Missing", "Medium Missing",
               "Last Scan", "Days Since Patch", "Compliant", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    for data_row in range(4, 24):
        for col in range(1, 10):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15


def create_byod_assessment_sheet(ws, styles):
    ws.merge_cells("A1:L1")
    ws["A1"] = "BYOD Security Assessment"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    headers = ["Device ID", "Device Type", "Owner", "OS/Version", "Min OS Met", "MDM Enrolled",
               "Container", "Encrypted", "Remote Wipe", "Agreement Signed", "Compliant", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    for data_row in range(4, 19):
        for col in range(1, 13):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 13


def create_physical_security_sheet(ws, styles):
    ws.merge_cells("A1:I1")
    ws["A1"] = "Physical Security Requirements Assessment"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    headers = ["Category", "Requirement", "Level", "Documented", "Communicated",
               "Ack Required", "Ack Captured", "Verification", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row = 4
    for category, requirement, level in PHYSICAL_REQUIREMENTS:
        ws.cell(row=row, column=1, value=category).border = styles["border"]
        ws.cell(row=row, column=2, value=requirement).border = styles["border"]
        ws.cell(row=row, column=3, value=level).border = styles["border"]
        for col in range(4, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 45
    for col in range(3, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14


def create_lost_stolen_sheet(ws, styles):
    ws.merge_cells("A1:I1")
    ws["A1"] = "Lost/Stolen Device Procedures Assessment"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    headers = ["Procedure Element", "Requirement", "SLA", "Implemented", "Documentation",
               "Last Tested", "Responsible", "Evidence", "Compliant"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row = 4
    for element, requirement, sla in LOST_STOLEN_PROCEDURES:
        ws.cell(row=row, column=1, value=element).border = styles["border"]
        ws.cell(row=row, column=2, value=requirement).border = styles["border"]
        ws.cell(row=row, column=3, value=sla).border = styles["border"]
        for col in range(4, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    for col in range(3, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15


def create_gap_analysis_sheet(ws, styles):
    ws.merge_cells("A1:I1")
    ws["A1"] = "Endpoint and Physical Security Gap Analysis"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    headers = ["Gap ID", "Source", "Description", "Scope", "Risk", "Remediation", "Owner", "Target", "Status"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    for data_row in range(4, 19):
        ws.cell(row=data_row, column=1, value=f"GAP-EPS-{data_row-3:03d}").border = styles["border"]
        for col in range(2, 10):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15


def create_evidence_register_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "Evidence Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    headers = ["Evidence ID", "Type", "Description", "Source", "Date", "Retention", "Location", "Status"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    evidence = [("EVD-EPS-001", "MDM Reports", "Device compliance reports", "MDM"),
                ("EVD-EPS-002", "EDR Reports", "Endpoint protection status", "EDR Console"),
                ("EVD-EPS-003", "Patch Reports", "Patch compliance dashboards", "WSUS/SCCM"),
                ("EVD-EPS-004", "Inventory Export", "Device inventory from MDM", "MDM"),
                ("EVD-EPS-005", "BYOD Policy", "BYOD policy documentation", "Policy Repo"),
                ("EVD-EPS-006", "Acknowledgments", "Physical security acknowledgments", "HR")]
    row = 4
    for evd_id, evd_type, desc, source in evidence:
        ws.cell(row=row, column=1, value=evd_id).border = styles["border"]
        ws.cell(row=row, column=2, value=evd_type).border = styles["border"]
        ws.cell(row=row, column=3, value=desc).border = styles["border"]
        ws.cell(row=row, column=4, value=source).border = styles["border"]
        for col in range(5, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18


def create_dashboard_sheet(ws, styles):
    ws.merge_cells("A1:D1")
    ws["A1"] = "Endpoint and Physical Security Dashboard"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    metrics = [("Total Devices in Inventory", ""), ("Encryption Compliance Rate", ""),
               ("Endpoint Protection Coverage", ""), ("Patch Compliance Rate", ""),
               ("BYOD Compliance Rate", ""), ("Physical Security Acknowledgment Rate", ""),
               ("Open Gaps (High/Critical)", "")]
    row = 4
    for metric, value in metrics:
        ws[f"A{row}"] = metric
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = styles["border"]
        cell = ws[f"B{row}"]
        cell.fill = styles["input_cell"]["fill"]
        cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20


def create_approval_sheet(ws, styles):
    ws.merge_cells("A1:E1")
    ws["A1"] = "Assessment Approval and Sign-Off"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    headers = ["Role", "Name", "Signature", "Date", "Comments"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    approvers = ["Endpoint Management Lead", "IT Security Manager", "Desktop Support Manager", "CISO"]
    row = 4
    for approver in approvers:
        ws.cell(row=row, column=1, value=approver).border = styles["border"]
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 40


def main():
    logger.info(f"Starting generation of {DOCUMENT_ID} - {WORKBOOK_NAME}")
    logger.info(f"Output file: {OUTPUT_FILENAME}")

    try:
        styles = setup_styles()
        wb = create_workbook()

        logger.info("Creating sheets...")
        create_instructions_sheet(wb["Instructions"], styles)
        create_device_inventory_sheet(wb["Device_Inventory"], styles)
        create_encryption_status_sheet(wb["Encryption_Status"], styles)
        create_endpoint_protection_sheet(wb["Endpoint_Protection"], styles)
        create_patch_compliance_sheet(wb["Patch_Compliance"], styles)
        create_byod_assessment_sheet(wb["BYOD_Assessment"], styles)
        create_physical_security_sheet(wb["Physical_Security"], styles)
        create_lost_stolen_sheet(wb["Lost_Stolen_Procedures"], styles)
        create_gap_analysis_sheet(wb["Gap_Analysis"], styles)
        create_evidence_register_sheet(wb["Evidence_Register"], styles)
        create_dashboard_sheet(wb["Dashboard"], styles)
        create_approval_sheet(wb["Approval_Sign_Off"], styles)

        wb.save(OUTPUT_FILENAME)
        logger.info(f"Workbook saved successfully: {OUTPUT_FILENAME}")

    except Exception as e:
        logger.error(f"Error generating workbook: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation, constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
