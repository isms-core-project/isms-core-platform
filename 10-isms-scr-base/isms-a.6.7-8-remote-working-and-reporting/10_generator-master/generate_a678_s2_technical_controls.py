#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.7-8.S2 - Technical Controls Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.6.7 (Remote Working) & A.6.8 (Event Reporting)
Assessment Domain S2 of S5: Technical Controls Assessment

Reference Pattern: Based on ISMS-IMP-A.6.7-8.S2 specification
================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.7-8.S2"
WORKBOOK_NAME = "Technical Controls Assessment"
CONTROL_ID = "A.6.7-8"
CONTROL_NAME = "Remote Working and Security Event Reporting"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header": {"font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
                   "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
                   "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)},
        "subheader": {"font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
                      "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
                      "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)},
        "column_header": {"font": Font(name="Calibri", size=10, bold=True),
                          "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
                          "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
                          "border": border_thin},
        "input_cell": {"fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
                       "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
                       "border": border_thin},
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }

# =============================================================================
# PRE-POPULATED DATA
# =============================================================================
VPN_REQUIREMENTS = [
    ("VPN Solution", "Enterprise VPN solution deployed for remote access", "Mandatory"),
    ("Split Tunneling", "Split tunneling disabled or controlled", "Mandatory"),
    ("Protocol Security", "Strong encryption protocols (IKEv2, OpenVPN, WireGuard)", "Mandatory"),
    ("Certificate Auth", "Certificate-based authentication where feasible", "Recommended"),
    ("Connection Logging", "VPN connection logging enabled", "Mandatory"),
    ("Timeout Settings", "Idle session timeout configured", "Mandatory"),
]

MFA_REQUIREMENTS = [
    ("MFA Deployed", "Multi-factor authentication required for all remote access", "Mandatory"),
    ("MFA Methods", "Approved MFA methods defined (authenticator app, hardware token)", "Mandatory"),
    ("Backup Methods", "Backup authentication methods available", "Recommended"),
    ("MFA for VPN", "MFA required for VPN connection", "Mandatory"),
    ("MFA for Apps", "MFA required for cloud/SaaS applications", "Mandatory"),
    ("Phishing-Resistant", "Phishing-resistant MFA considered for high-risk access", "Recommended"),
]

ENCRYPTION_REQUIREMENTS = [
    ("Data in Transit", "All remote connections use TLS 1.2+ minimum", "Mandatory"),
    ("Data at Rest", "Full-disk encryption on all remote devices", "Mandatory"),
    ("Key Management", "Encryption keys managed centrally", "Mandatory"),
    ("Recovery Keys", "Recovery keys escrowed for organization access", "Mandatory"),
    ("Email Encryption", "Email encryption available for sensitive communications", "Recommended"),
]

LOGGING_REQUIREMENTS = [
    ("Authentication Logs", "All remote authentication attempts logged", "Mandatory"),
    ("VPN Logs", "VPN connection/disconnection events captured", "Mandatory"),
    ("Access Logs", "Remote access to sensitive resources logged", "Mandatory"),
    ("Log Retention", "Logs retained per policy requirements", "Mandatory"),
    ("Log Monitoring", "Remote access logs included in security monitoring", "Mandatory"),
    ("Anomaly Detection", "Anomalous remote access patterns detected", "Recommended"),
]


def create_workbook() -> Workbook:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    sheets = ["Instructions", "VPN_Assessment", "MFA_Assessment", "Encryption_Assessment",
              "Logging_Assessment", "Compliance_Summary", "Gap_Analysis", "Evidence_Register",
              "Dashboard", "Approval_Sign_Off"]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def create_instructions_sheet(ws, styles):
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    doc_info = [("Document ID", DOCUMENT_ID), ("Assessment Area", "VPN, MFA, Encryption, Logging"),
                ("Related Policy", "ISMS-POL-A.6.7-8, Section 2.3 (Technical Security)"),
                ("Version", "1.0"), ("Assessment Date", ""), ("Completed By", "")]
    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


def create_requirement_sheet(ws, styles, title, requirements, sheet_name):
    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Requirement", "Description", "Level", "Implemented", "Evidence", "Compliance", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row = 4
    for req, desc, level in requirements:
        ws.cell(row=row, column=1, value=req).border = styles["border"]
        ws.cell(row=row, column=2, value=desc).border = styles["border"]
        ws.cell(row=row, column=3, value=level).border = styles["border"]
        for col in range(4, 8):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    dv = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"D4:D{row-1}")

    comp_dv = DataValidation(type="list", formula1='"Compliant,Non-Compliant,Partial"', allow_blank=True)
    ws.add_data_validation(comp_dv)
    comp_dv.add(f"F4:F{row-1}")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 30


def create_compliance_summary_sheet(ws, styles):
    ws.merge_cells("A1:E1")
    ws["A1"] = "Technical Controls Compliance Summary"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Control Area", "Total Requirements", "Compliant", "Non-Compliant", "Compliance %"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    areas = [("VPN Controls", len(VPN_REQUIREMENTS)), ("MFA Controls", len(MFA_REQUIREMENTS)),
             ("Encryption Controls", len(ENCRYPTION_REQUIREMENTS)), ("Logging Controls", len(LOGGING_REQUIREMENTS))]
    row = 4
    for area, count in areas:
        ws.cell(row=row, column=1, value=area).border = styles["border"]
        ws.cell(row=row, column=2, value=count).border = styles["border"]
        for col in range(3, 6):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15


def create_gap_analysis_sheet(ws, styles):
    ws.merge_cells("A1:I1")
    ws["A1"] = "Technical Controls Gap Analysis"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Gap ID", "Control Area", "Gap Description", "Impact", "Risk Level",
               "Remediation", "Owner", "Target Date", "Status"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    for data_row in range(4, 19):
        ws.cell(row=data_row, column=1, value=f"GAP-TC-{data_row-3:03d}").border = styles["border"]
        for col in range(2, 10):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    area_dv = DataValidation(type="list", formula1='"VPN,MFA,Encryption,Logging"', allow_blank=True)
    ws.add_data_validation(area_dv)
    area_dv.add("B4:B18")

    risk_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add("E4:E18")

    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Accepted"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("I4:I18")

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15


def create_evidence_register_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "Technical Controls Evidence Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Evidence ID", "Evidence Type", "Description", "Source", "Date", "Retention", "Location", "Status"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    evidence = [("EVD-TC-001", "VPN Configuration", "VPN server configuration documentation", "IT Ops"),
                ("EVD-TC-002", "VPN Logs", "Sample VPN connection logs", "SIEM"),
                ("EVD-TC-003", "MFA Enrollment", "MFA enrollment statistics", "IAM System"),
                ("EVD-TC-004", "Encryption Status", "Device encryption compliance report", "MDM"),
                ("EVD-TC-005", "Log Retention Policy", "Log retention configuration", "SIEM")]
    row = 4
    for evd_id, evd_type, desc, source in evidence:
        ws.cell(row=row, column=1, value=evd_id).border = styles["border"]
        ws.cell(row=row, column=2, value=evd_type).border = styles["border"]
        ws.cell(row=row, column=3, value=desc).border = styles["border"]
        ws.cell(row=row, column=4, value=source).border = styles["border"]
        for col in range(5, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18


def create_dashboard_sheet(ws, styles):
    ws.merge_cells("A1:D1")
    ws["A1"] = "Technical Controls Dashboard"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    metrics = [("VPN Compliance Rate", ""), ("MFA Enrollment Rate", ""), ("Device Encryption Rate", ""),
               ("Logging Coverage", ""), ("Open Technical Gaps", ""), ("Overall Technical Compliance", "")]
    row = 4
    for metric, value in metrics:
        ws[f"A{row}"] = metric
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = styles["border"]
        cell = ws[f"B{row}"]
        cell.fill = styles["input_cell"]["fill"]
        cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20


def create_approval_sheet(ws, styles):
    ws.merge_cells("A1:E1")
    ws["A1"] = "Assessment Approval and Sign-Off"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Role", "Name", "Signature", "Date", "Comments"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    approvers = ["IT Security Manager", "Network Operations Manager", "IT Director", "CISO"]
    row = 4
    for approver in approvers:
        ws.cell(row=row, column=1, value=approver).border = styles["border"]
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 40


def main():
    logger.info(f"Starting generation of {DOCUMENT_ID} - {WORKBOOK_NAME}")
    logger.info(f"Output file: {OUTPUT_FILENAME}")

    try:
        styles = setup_styles()
        wb = create_workbook()

        logger.info("Creating sheets...")
        create_instructions_sheet(wb["Instructions"], styles)
        create_requirement_sheet(wb["VPN_Assessment"], styles, "VPN Security Assessment", VPN_REQUIREMENTS, "VPN")
        create_requirement_sheet(wb["MFA_Assessment"], styles, "MFA Security Assessment", MFA_REQUIREMENTS, "MFA")
        create_requirement_sheet(wb["Encryption_Assessment"], styles, "Encryption Assessment", ENCRYPTION_REQUIREMENTS, "Encryption")
        create_requirement_sheet(wb["Logging_Assessment"], styles, "Logging Assessment", LOGGING_REQUIREMENTS, "Logging")
        create_compliance_summary_sheet(wb["Compliance_Summary"], styles)
        create_gap_analysis_sheet(wb["Gap_Analysis"], styles)
        create_evidence_register_sheet(wb["Evidence_Register"], styles)
        create_dashboard_sheet(wb["Dashboard"], styles)
        create_approval_sheet(wb["Approval_Sign_Off"], styles)

        wb.save(OUTPUT_FILENAME)
        logger.info(f"Workbook saved successfully: {OUTPUT_FILENAME}")

    except Exception as e:
        logger.error(f"Error generating workbook: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation, constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
