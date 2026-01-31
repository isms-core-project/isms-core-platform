#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.33-34.1 - Test Data Protection Assessment
================================================================================

ISO/IEC 27001:2022 Controls A.8.33: Test Information & A.8.34: Protection of
Information Systems During Audit Testing
Assessment Domain 1 of 3: Test Data Governance & Protection

This script generates a comprehensive Excel assessment workbook for managing
test data protection including inventory, masking, anonymization, environment
registry, and compliance verification.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.33-34.1"
WORKBOOK_NAME = "Test Data Protection Assessment"
CONTROL_ID = "A.8.33-34"
CONTROL_NAME = "Testing and Audit Protection"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions and Legend sheet."""
    ws.title = "Instructions_and_Legend"

    instructions = [
        ["ISMS-IMP-A.8.33-34.1 - Test Data Protection Assessment"],
        ["ISO/IEC 27001:2022 - Controls A.8.33 & A.8.34: Test Data Governance"],
        [""],
        ["PURPOSE"],
        ["This workbook tracks test data governance and protection compliance including"],
        ["test data inventory, masking effectiveness, environment security, and refresh governance."],
        [""],
        ["SHEETS"],
        ["1. Instructions_and_Legend - This guidance sheet"],
        ["2. Test_Data_Inventory - Registry of all production data copied to test environments"],
        ["3. Data_Masking_Assessment - Evaluation of data masking and anonymization effectiveness"],
        ["4. Test_Environment_Registry - Inventory of all test environments with security posture"],
        ["5. Data_Refresh_Schedule - Governance of test data refresh cycles and authorizations"],
        ["6. Compliance_Verification - Documentation of compliance status and verification activities"],
        ["7. Summary_Dashboard - Executive overview of test data governance compliance"],
        ["8. Evidence_Register - Documentation supporting assessment findings"],
        ["9. Approval_Sign_Off - Three-level approval workflow"],
        [""],
        ["STATUS LEGEND"],
        ["Compliant - Fully meets requirement"],
        ["Partial - Partially meets requirement, improvement needed"],
        ["Non-Compliant - Does not meet requirement"],
        ["Planned - Implementation planned"],
        ["N/A - Not Applicable"],
        [""],
        ["MASKING EFFECTIVENESS SCALE"],
        ["5 - Excellent: All PII fully anonymized, irreversible"],
        ["4 - Good: Strong masking, minimal re-identification risk"],
        ["3 - Adequate: Reasonable masking, some risk"],
        ["2 - Poor: Weak masking, significant risk"],
        ["1 - Inadequate: Minimal/no masking, high risk"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14, color="003366")
            elif row_num == 2:
                cell.font = Font(bold=True, size=11, color="003366")
            elif value in ["PURPOSE", "SHEETS", "STATUS LEGEND", "MASKING EFFECTIVENESS SCALE"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 80


def create_test_data_inventory_sheet(ws):
    """Create the Test Data Inventory sheet."""
    ws.title = "Test_Data_Inventory"

    # Title
    ws.merge_cells('A1:T1')
    title_cell = ws.cell(row=1, column=1, value="TEST DATA INVENTORY")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Registry of all production data copied to test environments")

    headers = [
        "Data_Set_ID", "Data_Set_Name", "Source_System", "Target_Environment",
        "Data_Classification", "Contains_PII", "PII_Categories", "Data_Volume",
        "Authorization_Status", "Data_Owner", "Authorizer", "Authorization_Date",
        "Last_Copy_Date", "Refresh_Frequency", "Masking_Required", "Masking_Status",
        "Business_Justification", "Expiration_Date", "Evidence_Reference", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    classification_dv = DataValidation(type="list", formula1='"Public,Internal,Confidential,Restricted"')
    ws.add_data_validation(classification_dv)
    classification_dv.add('E5:E104')

    pii_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(pii_dv)
    pii_dv.add('F5:F104')
    pii_dv.add('O5:O104')

    auth_status_dv = DataValidation(type="list", formula1='"Authorized,Pending,Unauthorized"')
    ws.add_data_validation(auth_status_dv)
    auth_status_dv.add('I5:I104')

    refresh_dv = DataValidation(type="list", formula1='"Daily,Weekly,Monthly,Quarterly,Ad-Hoc,One-Time"')
    ws.add_data_validation(refresh_dv)
    refresh_dv.add('N5:N104')

    masking_dv = DataValidation(type="list", formula1='"Fully Masked,Partially Masked,Not Masked,N/A"')
    ws.add_data_validation(masking_dv)
    masking_dv.add('P5:P104')

    # Format input rows
    for row in range(5, 105):
        ws.cell(row=row, column=1, value=f"TDI-{str(row-4).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics section
    ws.cell(row=107, column=1, value="SUMMARY STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Data Sets", "=COUNTA(B5:B104)"),
        ("Authorized Data Sets", '=COUNTIF(I5:I104,"Authorized")'),
        ("Pending Authorization", '=COUNTIF(I5:I104,"Pending")'),
        ("Unauthorized Data Sets", '=COUNTIF(I5:I104,"Unauthorized")'),
        ("Data Sets with PII", '=COUNTIF(F5:F104,"Yes")'),
        ("Fully Masked PII Sets", '=COUNTIFS(F5:F104,"Yes",P5:P104,"Fully Masked")'),
        ("Partially Masked PII Sets", '=COUNTIFS(F5:F104,"Yes",P5:P104,"Partially Masked")'),
        ("Unmasked PII Sets", '=COUNTIFS(F5:F104,"Yes",P5:P104,"Not Masked")'),
    ]

    row = 109
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [15, 30, 25, 25, 18, 12, 30, 15, 18, 25, 25, 15, 15, 18, 12, 18, 40, 15, 20, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_data_masking_assessment_sheet(ws):
    """Create the Data Masking Assessment sheet."""
    ws.title = "Data_Masking_Assessment"

    ws.merge_cells('A1:U1')
    title_cell = ws.cell(row=1, column=1, value="DATA MASKING ASSESSMENT")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Evaluation of data masking and anonymization effectiveness")

    headers = [
        "Data_Set_ID", "Data_Set_Name", "Target_Environment", "Contains_PII",
        "Masking_Status", "Primary_Masking_Technique", "Masking_Tool",
        "Masking_Effectiveness_Score", "PII_Fields_Identified", "PII_Fields_Masked",
        "PII_Fields_Unmasked", "Masking_Verification_Date", "Verification_Method",
        "Re_identification_Risk", "Masking_Gap_Severity", "Remediation_Owner",
        "Remediation_Target_Date", "Remediation_Status", "Exception_Approved",
        "Exception_Justification", "Evidence_Reference"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    pii_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(pii_dv)
    pii_dv.add('D5:D104')
    pii_dv.add('S5:S104')

    masking_dv = DataValidation(type="list", formula1='"Fully Masked,Partially Masked,Not Masked,N/A"')
    ws.add_data_validation(masking_dv)
    masking_dv.add('E5:E104')

    technique_dv = DataValidation(type="list", formula1='"Substitution,Shuffling,Tokenization,Encryption,Synthetic,Anonymization,None"')
    ws.add_data_validation(technique_dv)
    technique_dv.add('F5:F104')

    score_dv = DataValidation(type="list", formula1='"1,2,3,4,5"')
    ws.add_data_validation(score_dv)
    score_dv.add('H5:H104')

    verify_dv = DataValidation(type="list", formula1='"Automated,Manual Sampling,None"')
    ws.add_data_validation(verify_dv)
    verify_dv.add('M5:M104')

    risk_dv = DataValidation(type="list", formula1='"High,Medium,Low,None"')
    ws.add_data_validation(risk_dv)
    risk_dv.add('N5:N104')

    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(severity_dv)
    severity_dv.add('O5:O104')

    status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,Completed"')
    ws.add_data_validation(status_dv)
    status_dv.add('R5:R104')

    # Format input rows
    for row in range(5, 105):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics
    ws.cell(row=107, column=1, value="MASKING TECHNIQUE SUMMARY").font = Font(bold=True, size=12)

    techniques = ["Substitution", "Shuffling", "Tokenization", "Encryption", "Synthetic", "Anonymization", "None"]
    row = 109
    for tech in techniques:
        ws.cell(row=row, column=1, value=tech).border = THIN_BORDER
        ws.cell(row=row, column=2, value=f'=COUNTIF(F5:F104,"{tech}")').border = THIN_BORDER
        row += 1

    ws.cell(row=row + 1, column=1, value="EFFECTIVENESS SUMMARY").font = Font(bold=True, size=12)
    row += 3
    ws.cell(row=row, column=1, value="Average Effectiveness Score").border = THIN_BORDER
    ws.cell(row=row, column=2, value="=AVERAGE(H5:H104)").border = THIN_BORDER
    row += 1
    ws.cell(row=row, column=1, value="High Re-identification Risk").border = THIN_BORDER
    ws.cell(row=row, column=2, value='=COUNTIF(N5:N104,"High")').border = THIN_BORDER
    row += 1
    ws.cell(row=row, column=1, value="Critical Masking Gaps").border = THIN_BORDER
    ws.cell(row=row, column=2, value='=COUNTIF(O5:O104,"Critical")').border = THIN_BORDER

    # Column widths
    widths = [15, 30, 25, 12, 18, 25, 25, 12, 40, 40, 40, 15, 25, 18, 18, 25, 15, 18, 12, 40, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_test_environment_registry_sheet(ws):
    """Create the Test Environment Registry sheet."""
    ws.title = "Test_Environment_Registry"

    ws.merge_cells('A1:U1')
    title_cell = ws.cell(row=1, column=1, value="TEST ENVIRONMENT REGISTRY")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Inventory of all test environments with security posture assessment")

    headers = [
        "Environment_ID", "Environment_Name", "Environment_Type", "Infrastructure_Type",
        "Environment_Owner", "Business_Unit", "Highest_Data_Classification",
        "Contains_Production_Data", "Access_Control_Type", "Network_Isolation",
        "Encryption_at_Rest", "Encryption_in_Transit", "Logging_Enabled",
        "Patch_Management", "Security_Control_Status", "Last_Security_Review",
        "Next_Review_Due", "Data_Masking_Enforced", "Environment_URL_Location",
        "Support_Contact", "Evidence_Reference"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    env_type_dv = DataValidation(type="list", formula1='"Development,QA,Staging,UAT,Performance,Training,DR-Test,Sandbox"')
    ws.add_data_validation(env_type_dv)
    env_type_dv.add('C5:C54')

    infra_type_dv = DataValidation(type="list", formula1='"On-Premise,Cloud-AWS,Cloud-Azure,Cloud-GCP,Hybrid,Container,Local"')
    ws.add_data_validation(infra_type_dv)
    infra_type_dv.add('D5:D54')

    classification_dv = DataValidation(type="list", formula1='"Public,Internal,Confidential,Restricted"')
    ws.add_data_validation(classification_dv)
    classification_dv.add('G5:G54')

    yn_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('H5:H54')

    access_dv = DataValidation(type="list", formula1='"RBAC,AD/LDAP,SSO,Local Accounts,None"')
    ws.add_data_validation(access_dv)
    access_dv.add('I5:I54')

    isolation_dv = DataValidation(type="list", formula1='"Full,Partial,None"')
    ws.add_data_validation(isolation_dv)
    isolation_dv.add('J5:J54')

    ynp_dv = DataValidation(type="list", formula1='"Yes,Partial,No"')
    ws.add_data_validation(ynp_dv)
    ynp_dv.add('K5:K54')
    ynp_dv.add('L5:L54')
    ynp_dv.add('M5:M54')
    ynp_dv.add('R5:R54')

    patch_dv = DataValidation(type="list", formula1='"Automated,Manual-Current,Manual-Delayed,None"')
    ws.add_data_validation(patch_dv)
    patch_dv.add('N5:N54')

    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant"')
    ws.add_data_validation(status_dv)
    status_dv.add('O5:O54')

    # Format input rows
    for row in range(5, 55):
        ws.cell(row=row, column=1, value=f"ENV-{str(row-4).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics
    ws.cell(row=57, column=1, value="ENVIRONMENT STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Environments", "=COUNTA(B5:B54)"),
        ("Development Environments", '=COUNTIF(C5:C54,"Development")'),
        ("QA Environments", '=COUNTIF(C5:C54,"QA")'),
        ("UAT Environments", '=COUNTIF(C5:C54,"UAT")'),
        ("Environments with Prod Data", '=COUNTIF(H5:H54,"Yes")'),
        ("Security Compliant", '=COUNTIF(O5:O54,"Compliant")'),
        ("Security Partial", '=COUNTIF(O5:O54,"Partial")'),
        ("Security Non-Compliant", '=COUNTIF(O5:O54,"Non-Compliant")'),
        ("Full Network Isolation", '=COUNTIF(J5:J54,"Full")'),
        ("No Network Isolation", '=COUNTIF(J5:J54,"None")'),
    ]

    row = 59
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [15, 30, 20, 20, 25, 25, 18, 12, 25, 18, 12, 12, 12, 18, 18, 15, 15, 12, 40, 25, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_data_refresh_schedule_sheet(ws):
    """Create the Data Refresh Schedule sheet."""
    ws.title = "Data_Refresh_Schedule"

    ws.merge_cells('A1:T1')
    title_cell = ws.cell(row=1, column=1, value="DATA REFRESH SCHEDULE")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Governance of test data refresh cycles and authorizations")

    headers = [
        "Refresh_ID", "Target_Environment", "Data_Sources", "Refresh_Frequency",
        "Refresh_Method", "Last_Refresh_Date", "Next_Scheduled_Refresh",
        "Authorization_Status", "Authorizer", "Authorization_Date",
        "Masking_Applied_at_Refresh", "Masking_Tool", "Data_Volume",
        "Refresh_Duration", "Refresh_Window", "Retention_Period",
        "Auto_Purge_Enabled", "Refresh_Owner", "Refresh_Log_Location", "Evidence_Reference"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    freq_dv = DataValidation(type="list", formula1='"Daily,Weekly,Bi-Weekly,Monthly,Quarterly,Ad-Hoc"')
    ws.add_data_validation(freq_dv)
    freq_dv.add('D5:D54')

    method_dv = DataValidation(type="list", formula1='"Full Copy,Incremental,Subset,Synthetic,Clone"')
    ws.add_data_validation(method_dv)
    method_dv.add('E5:E54')

    auth_dv = DataValidation(type="list", formula1='"Authorized,Pending,Unauthorized"')
    ws.add_data_validation(auth_dv)
    auth_dv.add('H5:H54')

    masking_dv = DataValidation(type="list", formula1='"Yes - Automated,Yes - Manual,Partial,No"')
    ws.add_data_validation(masking_dv)
    masking_dv.add('K5:K54')

    yn_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('Q5:Q54')

    # Format input rows
    for row in range(5, 55):
        ws.cell(row=row, column=1, value=f"REF-{str(row-4).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics
    ws.cell(row=57, column=1, value="REFRESH STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Refresh Schedules", "=COUNTA(B5:B54)"),
        ("Authorized Refreshes", '=COUNTIF(H5:H54,"Authorized")'),
        ("Unauthorized Refreshes", '=COUNTIF(H5:H54,"Unauthorized")'),
        ("Daily Refreshes", '=COUNTIF(D5:D54,"Daily")'),
        ("Weekly Refreshes", '=COUNTIF(D5:D54,"Weekly")'),
        ("Masking Automated", '=COUNTIF(K5:K54,"Yes - Automated")'),
        ("No Masking at Refresh", '=COUNTIF(K5:K54,"No")'),
        ("Auto-Purge Enabled", '=COUNTIF(Q5:Q54,"Yes")'),
    ]

    row = 59
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [15, 25, 40, 18, 25, 15, 15, 18, 25, 15, 18, 25, 15, 15, 20, 18, 12, 25, 30, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_compliance_verification_sheet(ws):
    """Create the Compliance Verification sheet."""
    ws.title = "Compliance_Verification"

    ws.merge_cells('A1:R1')
    title_cell = ws.cell(row=1, column=1, value="COMPLIANCE VERIFICATION")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Documentation of compliance status and verification activities")

    headers = [
        "Requirement_ID", "Requirement_Source", "Requirement_Reference",
        "Requirement_Description", "Applicable_Data_Sets", "Applicable_Environments",
        "Compliance_Status", "Last_Verification_Date", "Verification_Method",
        "Verifier", "Findings", "Finding_Severity", "Remediation_Required",
        "Remediation_Owner", "Remediation_Target_Date", "Remediation_Status",
        "Next_Verification_Due", "Evidence_Reference"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    source_dv = DataValidation(type="list", formula1='"GDPR,ISO 27001,FADP,Internal Policy,Industry Standard"')
    ws.add_data_validation(source_dv)
    source_dv.add('B5:B54')

    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant,Not Assessed"')
    ws.add_data_validation(status_dv)
    status_dv.add('G5:G54')

    method_dv = DataValidation(type="list", formula1='"Automated Check,Manual Audit,Self-Assessment,Third-Party Audit"')
    ws.add_data_validation(method_dv)
    method_dv.add('I5:I54')

    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(severity_dv)
    severity_dv.add('L5:L54')

    yn_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('M5:M54')

    rem_status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,Completed"')
    ws.add_data_validation(rem_status_dv)
    rem_status_dv.add('P5:P54')

    # Pre-populate standard requirements
    requirements = [
        ("REQ-001", "GDPR", "Article 25", "Data protection by design - pseudonymization"),
        ("REQ-002", "GDPR", "Article 32", "Security of processing - technical measures"),
        ("REQ-003", "ISO 27001", "A.8.33", "Test information appropriately protected"),
        ("REQ-004", "ISO 27001", "A.8.34", "Audit testing planned and agreed"),
        ("REQ-005", "Internal Policy", "POL-2.1", "All production data copies authorized"),
        ("REQ-006", "Internal Policy", "POL-2.2", "PII must be masked in test environments"),
        ("REQ-007", "Internal Policy", "POL-2.3", "Test environments must have security controls"),
        ("REQ-008", "Internal Policy", "POL-2.4", "Data refresh must be authorized"),
        ("REQ-009", "FADP", "Article 8", "Data security requirements"),
        ("REQ-010", "Industry Standard", "PCI-DSS 6.4", "Test data security (if applicable)"),
    ]

    for row_idx, (req_id, source, ref, desc) in enumerate(requirements, 5):
        ws.cell(row=row_idx, column=1, value=req_id)
        ws.cell(row=row_idx, column=2, value=source)
        ws.cell(row=row_idx, column=3, value=ref)
        ws.cell(row=row_idx, column=4, value=desc)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Additional empty rows
    for row in range(15, 55):
        ws.cell(row=row, column=1, value=f"REQ-{str(row-4).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics
    ws.cell(row=57, column=1, value="COMPLIANCE STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Requirements", "=COUNTA(B5:B54)"),
        ("Compliant", '=COUNTIF(G5:G54,"Compliant")'),
        ("Partial", '=COUNTIF(G5:G54,"Partial")'),
        ("Non-Compliant", '=COUNTIF(G5:G54,"Non-Compliant")'),
        ("Not Assessed", '=COUNTIF(G5:G54,"Not Assessed")'),
        ("Critical Findings", '=COUNTIF(L5:L54,"Critical")'),
        ("High Findings", '=COUNTIF(L5:L54,"High")'),
    ]

    row = 59
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [15, 25, 25, 50, 40, 40, 18, 15, 25, 25, 50, 18, 12, 25, 15, 18, 15, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_summary_dashboard_sheet(ws):
    """Create the Summary Dashboard sheet."""
    ws.title = "Summary_Dashboard"

    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="TEST DATA PROTECTION - SUMMARY DASHBOARD")
    title_cell.font = Font(bold=True, size=16, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Executive overview of test data governance compliance")
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.cell(row=3, column=1, value=f"Report Date: {GENERATED_DATE}")

    # Overall Compliance Summary
    ws.cell(row=5, column=1, value="OVERALL COMPLIANCE SUMMARY").font = Font(bold=True, size=12)

    headers = ["Metric", "Value", "Target", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    metrics = [
        ("Data Set Authorization Rate", "[%]", "100%"),
        ("PII Masking Coverage", "[%]", "100%"),
        ("Average Masking Effectiveness", "[Score]", ">=4.0"),
        ("Environment Security Compliance", "[%]", ">=90%"),
        ("Refresh Authorization Rate", "[%]", "100%"),
        ("Regulatory Compliance Rate", "[%]", "100%"),
    ]

    row = 7
    for metric, value, target in metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        cell_val = ws.cell(row=row, column=2, value=value)
        cell_val.fill = INPUT_FILL
        cell_val.border = THIN_BORDER
        ws.cell(row=row, column=3, value=target).border = THIN_BORDER
        cell_status = ws.cell(row=row, column=4)
        cell_status.fill = INPUT_FILL
        cell_status.border = THIN_BORDER
        row += 1

    # Status validation
    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant"')
    ws.add_data_validation(status_dv)
    status_dv.add('D7:D15')

    # Critical Findings
    ws.cell(row=15, column=1, value="CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION").font = Font(bold=True, size=12, color="C00000")

    findings_headers = ["Category", "Finding", "Count", "Severity", "Action Required"]
    for col, header in enumerate(findings_headers, 1):
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    findings = [
        ("Unauthorized Data", "Data sets without authorization", "[#]", "Critical", "Immediate"),
        ("Unmasked PII", "PII data without masking", "[#]", "Critical", "Urgent"),
        ("Non-Compliant Environments", "Environments failing security", "[#]", "High", "Priority"),
        ("Unauthorized Refresh", "Refresh without approval", "[#]", "High", "Priority"),
        ("Overdue Reviews", "Security reviews past due", "[#]", "Medium", "Plan"),
    ]

    row = 17
    for cat, finding, count, severity, action in findings:
        ws.cell(row=row, column=1, value=cat).border = THIN_BORDER
        ws.cell(row=row, column=2, value=finding).border = THIN_BORDER
        cell_count = ws.cell(row=row, column=3, value=count)
        cell_count.fill = INPUT_FILL
        cell_count.border = THIN_BORDER
        ws.cell(row=row, column=4, value=severity).border = THIN_BORDER
        ws.cell(row=row, column=5, value=action).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [30, 40, 15, 15, 20, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_evidence_register_sheet(ws):
    """Create the Evidence Register sheet."""
    ws.title = "Evidence_Register"

    ws.merge_cells('A1:P1')
    title_cell = ws.cell(row=1, column=1, value="EVIDENCE REGISTER")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Documentation supporting assessment findings (100 entries)")

    headers = [
        "Evidence_ID", "Evidence_Type", "Evidence_Title", "Description",
        "Related_Assessment_Area", "Related_Finding_Control", "Document_Location",
        "Date_Collected", "Collected_By", "Verification_Status", "Verified_By",
        "Verification_Date", "Retention_Period", "Expiration_Date",
        "Confidentiality", "Auditor_Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(type="list", formula1='"Policy Document,Technical Config,Authorization Record,Audit Log,Verification Report,Training Record,Other"')
    ws.add_data_validation(type_dv)
    type_dv.add('B5:B104')

    area_dv = DataValidation(type="list", formula1='"Test Data Inventory,Masking,Environment,Refresh,Compliance"')
    ws.add_data_validation(area_dv)
    area_dv.add('E5:E104')

    status_dv = DataValidation(type="list", formula1='"Verified,Pending,Not Verified"')
    ws.add_data_validation(status_dv)
    status_dv.add('J5:J104')

    retention_dv = DataValidation(type="list", formula1='"1 Year,2 Years,3 Years,Permanent"')
    ws.add_data_validation(retention_dv)
    retention_dv.add('M5:M104')

    conf_dv = DataValidation(type="list", formula1='"Public,Internal,Confidential,Restricted"')
    ws.add_data_validation(conf_dv)
    conf_dv.add('O5:O104')

    # Format input rows
    for row in range(5, 105):
        ws.cell(row=row, column=1, value=f"EV-{str(row-4).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics
    ws.cell(row=107, column=1, value="EVIDENCE STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Evidence Entries", "=COUNTA(B5:B104)"),
        ("Verified Evidence", '=COUNTIF(J5:J104,"Verified")'),
        ("Pending Verification", '=COUNTIF(J5:J104,"Pending")'),
        ("Evidence by Area - Inventory", '=COUNTIF(E5:E104,"Test Data Inventory")'),
        ("Evidence by Area - Masking", '=COUNTIF(E5:E104,"Masking")'),
        ("Evidence by Area - Environment", '=COUNTIF(E5:E104,"Environment")'),
        ("Evidence by Area - Refresh", '=COUNTIF(E5:E104,"Refresh")'),
        ("Evidence by Area - Compliance", '=COUNTIF(E5:E104,"Compliance")'),
    ]

    row = 109
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [15, 25, 40, 50, 25, 25, 40, 15, 25, 18, 25, 15, 18, 15, 18, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_approval_sign_off_sheet(ws):
    """Create the Approval Sign-Off sheet."""
    ws.title = "Approval_Sign_Off"

    ws.merge_cells('A1:D1')
    title_cell = ws.cell(row=1, column=1, value="ASSESSMENT APPROVAL & SIGN-OFF")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Three-level approval workflow for test data protection assessment")

    # Assessment Summary
    ws.cell(row=4, column=1, value="ASSESSMENT SUMMARY").font = Font(bold=True, size=12)

    summary_items = [
        ("Document ID:", DOCUMENT_ID),
        ("Assessment Type:", "Test Data Protection Assessment"),
        ("Assessment Period:", "[USER INPUT]"),
        ("Overall Compliance:", "[Formula from Summary_Dashboard]"),
        ("Critical Findings:", "[Formula]"),
        ("High Findings:", "[Formula]"),
        ("Evidence Entries:", "[Formula]"),
        ("Assessment Status:", "[Dropdown]"),
    ]

    row = 6
    for label, value in summary_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value or "[Dropdown]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Status dropdown
    status_dv = DataValidation(type="list", formula1='"Final,Requires Action,Draft,Re-assessment Required"')
    ws.add_data_validation(status_dv)
    status_dv.add('B13')

    # Level 1: Technical Validation
    row += 2
    ws.cell(row=row, column=1, value="LEVEL 1: TECHNICAL VALIDATION").font = Font(bold=True, size=12)
    row += 2

    level1_items = [
        ("Validator Name:", "[USER INPUT]"),
        ("Role/Title:", "Test Manager / QA Lead"),
        ("Validation Date:", "[DD.MM.YYYY]"),
        ("Validation Notes:", "[Text area]"),
        ("Technical Accuracy:", "[Dropdown]"),
        ("Recommendation:", "[Dropdown]"),
        ("Signature:", "[USER INPUT]"),
    ]

    for label, value in level1_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value or "[DD.MM.YYYY]" in value or "[Text area]" in value or "[Dropdown]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Level 1 dropdowns
    tech_acc_dv = DataValidation(type="list", formula1='"Confirmed,Minor Issues,Major Issues"')
    ws.add_data_validation(tech_acc_dv)
    tech_acc_dv.add('B22')

    rec_dv = DataValidation(type="list", formula1='"Approve,Approve with Conditions,Reject"')
    ws.add_data_validation(rec_dv)
    rec_dv.add('B23')

    # Level 2: Security Approval
    row += 2
    ws.cell(row=row, column=1, value="LEVEL 2: SECURITY APPROVAL").font = Font(bold=True, size=12)
    row += 2

    level2_items = [
        ("Approver Name:", "[USER INPUT]"),
        ("Role/Title:", "CISO / Security Manager"),
        ("Approval Date:", "[DD.MM.YYYY]"),
        ("Security Assessment:", "[Dropdown]"),
        ("Risk Acceptance:", "[Dropdown]"),
        ("Approval Decision:", "[Dropdown]"),
        ("Conditions:", "[Text area]"),
        ("Signature:", "[USER INPUT]"),
    ]

    for label, value in level2_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value or "[DD.MM.YYYY]" in value or "[Text area]" in value or "[Dropdown]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Level 2 dropdowns
    sec_assess_dv = DataValidation(type="list", formula1='"Acceptable,Acceptable with Conditions,Unacceptable"')
    ws.add_data_validation(sec_assess_dv)
    sec_assess_dv.add('B32')

    risk_dv = DataValidation(type="list", formula1='"Yes - Risks Accepted,No - Remediation Required"')
    ws.add_data_validation(risk_dv)
    risk_dv.add('B33')

    app_dec_dv = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected"')
    ws.add_data_validation(app_dec_dv)
    app_dec_dv.add('B34')

    # Level 3: Compliance Confirmation
    row += 2
    ws.cell(row=row, column=1, value="LEVEL 3: COMPLIANCE CONFIRMATION").font = Font(bold=True, size=12)
    row += 2

    level3_items = [
        ("Approver Name:", "[USER INPUT]"),
        ("Role/Title:", "DPO / Compliance Officer"),
        ("Approval Date:", "[DD.MM.YYYY]"),
        ("Regulatory Compliance:", "[Dropdown]"),
        ("Approval Decision:", "[Dropdown]"),
        ("Conditions:", "[Text area]"),
        ("Signature:", "[USER INPUT]"),
    ]

    for label, value in level3_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value or "[DD.MM.YYYY]" in value or "[Text area]" in value or "[Dropdown]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Level 3 dropdowns
    reg_comp_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant"')
    ws.add_data_validation(reg_comp_dv)
    reg_comp_dv.add('B44')
    app_dec_dv.add('B45')

    # Next Review
    row += 2
    ws.cell(row=row, column=1, value="NEXT REVIEW").font = Font(bold=True, size=12)
    row += 2

    next_items = [
        ("Next Review Date:", "[Formula: Approval Date + 180 days]"),
        ("Review Responsibility:", "[USER INPUT]"),
        ("Focus Areas:", "[USER INPUT]"),
    ]

    for label, value in next_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30


def generate_workbook():
    """Generate the complete assessment workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Create all sheets
    create_instructions_sheet(wb.create_sheet())
    create_test_data_inventory_sheet(wb.create_sheet())
    create_data_masking_assessment_sheet(wb.create_sheet())
    create_test_environment_registry_sheet(wb.create_sheet())
    create_data_refresh_schedule_sheet(wb.create_sheet())
    create_compliance_verification_sheet(wb.create_sheet())
    create_summary_dashboard_sheet(wb.create_sheet())
    create_evidence_register_sheet(wb.create_sheet())
    create_approval_sign_off_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - INITIAL CREATION
# QA_TOOL: Claude Code
# CHANGES: Initial generator creation for A.8.33-34 Testing and Audit Protection
# =============================================================================
