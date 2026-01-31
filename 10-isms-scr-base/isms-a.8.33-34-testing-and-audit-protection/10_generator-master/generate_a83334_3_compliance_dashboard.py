#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.33-34.3 - Compliance Dashboard
================================================================================

ISO/IEC 27001:2022 Controls A.8.33: Test Information & A.8.34: Protection of
Information Systems During Audit Testing
Assessment Domain 3 of 3: Compliance Dashboard & Consolidated Oversight

This script generates a comprehensive Excel dashboard workbook that consolidates
all domain assessments into an executive compliance view with scoring, gap analysis,
exception tracking, and CISO certification.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.33-34.3"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.8.33-34"
CONTROL_NAME = "Testing and Audit Protection"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_executive_dashboard_sheet(ws):
    """Create the Executive Dashboard sheet."""
    ws.title = "Executive_Dashboard"

    ws.merge_cells('A1:G1')
    title_cell = ws.cell(row=1, column=1, value="ISMS-IMP-A.8.33-34.3 - Testing and Audit Protection Compliance Dashboard")
    title_cell.font = Font(bold=True, size=16, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="ISO/IEC 27001:2022 - Controls A.8.33 & A.8.34: Executive Overview")
    ws.cell(row=2, column=1).font = Font(bold=True, size=11, color="003366")

    # Document Information
    ws.cell(row=4, column=1, value="DOCUMENT INFORMATION").font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID:", DOCUMENT_ID),
        ("Report Type:", "Compliance Summary Dashboard"),
        ("Related Policy:", "ISMS-POL-A.8.33-34 (Test Information & Audit Protection)"),
        ("Version:", "1.0"),
        ("Report Date:", "[USER INPUT]"),
        ("Reporting Period:", "[USER INPUT]"),
        ("Prepared By:", "[USER INPUT]"),
        ("Organization:", "[Organization]"),
        ("Review Cycle:", "Quarterly"),
        ("Last Updated:", GENERATED_DATE),
    ]

    row = 5
    for label, value in doc_info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Overall Compliance Summary
    ws.cell(row=16, column=1, value="OVERALL COMPLIANCE SUMMARY").font = Font(bold=True, size=12)

    headers = ["Assessment Area", "Compliance %", "Status", "Critical Gaps", "Evidence Count", "Last Updated"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=17, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    assessments = [
        ("Test Data Protection (IMP-A.8.33-34.1)", "[%]", "[Status]", "[#]", "[#]", "[Date]"),
        ("Audit Activity Management (IMP-A.8.33-34.2)", "[%]", "[Status]", "[#]", "[#]", "[Date]"),
        ("OVERALL AVERAGE", "[%]", "[Status]", "[#]", "[#]", "-"),
    ]

    row = 18
    for assessment in assessments:
        for col, value in enumerate(assessment, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
            if row == 20:
                cell.font = Font(bold=True)
        row += 1

    # Status validation
    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant"')
    ws.add_data_validation(status_dv)
    status_dv.add('C18:C20')

    # Key Performance Indicators
    ws.cell(row=22, column=1, value="KEY PERFORMANCE INDICATORS").font = Font(bold=True, size=12)

    kpi_headers = ["KPI", "Target", "Current", "Status", "Trend", "Source"]
    for col, header in enumerate(kpi_headers, 1):
        cell = ws.cell(row=23, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    kpis = [
        ("TEST DATA DOMAIN", "", "", "", "", ""),
        ("Data Set Authorization Rate", "100%", "[%]", "[Status]", "[Trend]", "IMP-A.8.33-34.1"),
        ("PII Masking Coverage", "100%", "[%]", "[Status]", "[Trend]", "IMP-A.8.33-34.1"),
        ("Average Masking Effectiveness", ">=4.0", "[Score]", "[Status]", "[Trend]", "IMP-A.8.33-34.1"),
        ("Environment Security Compliance", ">=90%", "[%]", "[Status]", "[Trend]", "IMP-A.8.33-34.1"),
        ("Refresh Authorization Rate", "100%", "[%]", "[Status]", "[Trend]", "IMP-A.8.33-34.1"),
        ("AUDIT ACTIVITY DOMAIN", "", "", "", "", ""),
        ("Audit Authorization Rate", "100%", "[%]", "[Status]", "[Trend]", "IMP-A.8.33-34.2"),
        ("Tool Authorization Rate", "100%", "[%]", "[Status]", "[Trend]", "IMP-A.8.33-34.2"),
        ("Access Approval Rate", "100%", "[%]", "[Status]", "[Trend]", "IMP-A.8.33-34.2"),
        ("Access Revocation Timeliness", "100%", "[%]", "[Status]", "[Trend]", "IMP-A.8.33-34.2"),
        ("Critical System Mitigation Coverage", "100%", "[%]", "[Status]", "[Trend]", "IMP-A.8.33-34.2"),
        ("Evidence Encryption Rate", "100%", "[%]", "[Status]", "[Trend]", "IMP-A.8.33-34.2"),
    ]

    row = 24
    for kpi in kpis:
        for col, value in enumerate(kpi, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
            if "DOMAIN" in value:
                cell.font = Font(bold=True, color="003366")
        row += 1

    # Trend validation
    trend_dv = DataValidation(type="list", formula1='"Up,Stable,Down"')
    ws.add_data_validation(trend_dv)
    trend_dv.add('E25:E36')

    # Control Compliance Mapping
    ws.cell(row=39, column=1, value="ISO 27001:2022 CONTROL COMPLIANCE MAPPING").font = Font(bold=True, size=12)

    control_headers = ["ISO Control", "Requirement", "Assessment Source", "Compliance", "Status", "Evidence"]
    for col, header in enumerate(control_headers, 1):
        cell = ws.cell(row=40, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    controls = [
        ("A.8.33", "Test information appropriately protected", "IMP-A.8.33-34.1", "[%]", "[Status]", "[Link]"),
        ("A.8.33", "Production data copied only when authorized", "IMP-A.8.33-34.1", "[%]", "[Status]", "[Link]"),
        ("A.8.33", "Test data masked/anonymized", "IMP-A.8.33-34.1", "[%]", "[Status]", "[Link]"),
        ("A.8.33", "Test environments secured", "IMP-A.8.33-34.1", "[%]", "[Status]", "[Link]"),
        ("A.8.34", "Audit tests planned and agreed", "IMP-A.8.33-34.2", "[%]", "[Status]", "[Link]"),
        ("A.8.34", "Audit tools authorized", "IMP-A.8.33-34.2", "[%]", "[Status]", "[Link]"),
        ("A.8.34", "Auditor access controlled", "IMP-A.8.33-34.2", "[%]", "[Status]", "[Link]"),
        ("A.8.34", "Production systems protected during testing", "IMP-A.8.33-34.2", "[%]", "[Status]", "[Link]"),
        ("A.8.34", "Audit evidence protected", "IMP-A.8.33-34.2", "[%]", "[Status]", "[Link]"),
    ]

    row = 41
    for control in controls:
        for col, value in enumerate(control, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Critical Findings Summary
    ws.cell(row=52, column=1, value="CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION").font = Font(bold=True, size=12, color="C00000")

    findings_headers = ["Domain", "Critical Finding", "Severity", "Count", "Assigned To", "Target Date", "Status"]
    for col, header in enumerate(findings_headers, 1):
        cell = ws.cell(row=53, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    findings = [
        ("Test Data", "Unauthorized data in test", "Critical", "[#]", "[Name]", "[Date]", "[Status]"),
        ("Test Data", "Unmasked PII", "Critical", "[#]", "[Name]", "[Date]", "[Status]"),
        ("Audit", "Unauthorized audit activity", "Critical", "[#]", "[Name]", "[Date]", "[Status]"),
        ("Audit", "Unauthorized tools", "Critical", "[#]", "[Name]", "[Date]", "[Status]"),
        ("Audit", "Access not revoked", "Critical", "[#]", "[Name]", "[Date]", "[Status]"),
    ]

    row = 54
    for finding in findings:
        for col, value in enumerate(finding, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Column widths
    widths = [45, 20, 15, 15, 20, 15, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_test_data_compliance_sheet(ws):
    """Create the Test Data Compliance sheet."""
    ws.title = "Test_Data_Compliance"

    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="TEST DATA PROTECTION COMPLIANCE")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Detailed compliance status from IMP-A.8.33-34.1 Assessment")

    # Test Data Inventory Compliance
    ws.cell(row=4, column=1, value="TEST DATA INVENTORY COMPLIANCE").font = Font(bold=True, size=12)

    headers = ["Metric", "Target", "Current", "Status", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    inventory_metrics = [
        ("Total Data Sets", "-", "[#]", "-", ""),
        ("Authorized Data Sets", "100%", "[%]", "[Status]", ""),
        ("Pending Authorization", "0", "[#]", "[Status]", ""),
        ("Unauthorized Data Sets", "0", "[#]", "[Status]", "Critical if >0"),
        ("Data Sets with PII", "-", "[#]", "-", ""),
        ("PII Data Sets Masked", "100%", "[%]", "[Status]", ""),
        ("Authorization Rate %", "100%", "[%]", "[Status]", ""),
    ]

    row = 6
    for metric in inventory_metrics:
        for col, value in enumerate(metric, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Status validation
    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant,N/A"')
    ws.add_data_validation(status_dv)
    status_dv.add('D6:D50')

    # Masking Effectiveness Summary
    ws.cell(row=15, column=1, value="DATA MASKING EFFECTIVENESS").font = Font(bold=True, size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    masking_metrics = [
        ("Average Effectiveness Score", ">=4.0", "[Score]", "[Status]", ""),
        ("Score 5 (Excellent)", "-", "[#]", "-", ""),
        ("Score 4 (Good)", "-", "[#]", "-", ""),
        ("Score 3 (Adequate)", "-", "[#]", "-", ""),
        ("Score 2 (Poor)", "-", "[#]", "[Warning]", ""),
        ("Score 1 (Inadequate)", "-", "[#]", "[Critical]", ""),
        ("High Re-identification Risk", "0", "[#]", "[Status]", "Critical if >0"),
        ("Critical Masking Gaps", "0", "[#]", "[Status]", "Critical if >0"),
    ]

    row = 17
    for metric in masking_metrics:
        for col, value in enumerate(metric, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Environment Security Summary
    ws.cell(row=27, column=1, value="TEST ENVIRONMENT SECURITY").font = Font(bold=True, size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=28, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    env_metrics = [
        ("Total Environments", "-", "[#]", "-", ""),
        ("Security Compliant", "100%", "[%]", "[Status]", ""),
        ("Security Partial", "0%", "[%]", "[Status]", ""),
        ("Security Non-Compliant", "0", "[#]", "[Status]", ""),
        ("Full Network Isolation", ">=90%", "[%]", "[Status]", ""),
        ("Encryption at Rest", "100%", "[%]", "[Status]", ""),
        ("Reviews Overdue", "0", "[#]", "[Status]", ""),
    ]

    row = 29
    for metric in env_metrics:
        for col, value in enumerate(metric, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Refresh Governance Summary
    ws.cell(row=38, column=1, value="DATA REFRESH GOVERNANCE").font = Font(bold=True, size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=39, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    refresh_metrics = [
        ("Total Refresh Schedules", "-", "[#]", "-", ""),
        ("Authorized Refreshes", "100%", "[%]", "[Status]", ""),
        ("Masking at Refresh", "100%", "[%]", "[Status]", ""),
        ("Auto-Purge Enabled", "100%", "[%]", "[Status]", ""),
    ]

    row = 40
    for metric in refresh_metrics:
        for col, value in enumerate(metric, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Column widths
    widths = [35, 15, 15, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_audit_activity_compliance_sheet(ws):
    """Create the Audit Activity Compliance sheet."""
    ws.title = "Audit_Activity_Compliance"

    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="AUDIT ACTIVITY MANAGEMENT COMPLIANCE")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Detailed compliance status from IMP-A.8.33-34.2 Assessment")

    # Audit Activity Summary
    ws.cell(row=4, column=1, value="AUDIT ACTIVITY AUTHORIZATION").font = Font(bold=True, size=12)

    headers = ["Metric", "Target", "Current", "Status", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    audit_metrics = [
        ("Total Audits", "-", "[#]", "-", ""),
        ("Approved Audits", "100%", "[%]", "[Status]", ""),
        ("Pending Approval", "0", "[#]", "[Status]", ""),
        ("Completed Audits", "-", "[#]", "-", ""),
        ("Total Findings", "-", "[#]", "-", ""),
        ("Critical Findings", "0", "[#]", "[Status]", ""),
        ("Open Follow-ups", "-", "[#]", "-", "Track progress"),
    ]

    row = 6
    for metric in audit_metrics:
        for col, value in enumerate(metric, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Status validation
    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant,N/A"')
    ws.add_data_validation(status_dv)
    status_dv.add('D6:D70')

    # Tool Authorization Summary
    ws.cell(row=15, column=1, value="AUDIT TOOL AUTHORIZATION").font = Font(bold=True, size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    tool_metrics = [
        ("Total Tools", "-", "[#]", "-", ""),
        ("Authorized Tools", "100%", "[%]", "[Status]", ""),
        ("Pending Authorization", "0", "[#]", "[Status]", ""),
        ("Unauthorized Tools", "0", "[#]", "[Status]", "Critical"),
        ("High Risk Tools", "-", "[#]", "-", "Monitor closely"),
        ("Reviews Overdue", "0", "[#]", "[Status]", ""),
        ("Expired Licenses", "0", "[#]", "[Status]", ""),
    ]

    row = 17
    for metric in tool_metrics:
        for col, value in enumerate(metric, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Access Control Summary
    ws.cell(row=26, column=1, value="AUDITOR ACCESS CONTROL").font = Font(bold=True, size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=27, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    access_metrics = [
        ("Total Access Grants", "-", "[#]", "-", ""),
        ("Active Access", "-", "[#]", "-", "Monitor"),
        ("Approved Access", "100%", "[%]", "[Status]", ""),
        ("Pending Approval", "0", "[#]", "[Status]", ""),
        ("Admin Access Count", "Minimize", "[#]", "-", "High risk"),
        ("Restricted Data Access", "-", "[#]", "-", "Monitor"),
        ("NDA Coverage", "100%", "[%]", "[Status]", ""),
        ("Access Logging", "100%", "[%]", "[Status]", ""),
        ("MFA Required", "100%", "[%]", "[Status]", ""),
        ("Overdue Revocations", "0", "[#]", "[Status]", "Critical"),
    ]

    row = 28
    for metric in access_metrics:
        for col, value in enumerate(metric, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Disruption Mitigation Summary
    ws.cell(row=40, column=1, value="DISRUPTION MITIGATION COVERAGE").font = Font(bold=True, size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=41, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    mitigation_metrics = [
        ("Critical Systems", "-", "[#]", "-", ""),
        ("With Mitigation Plans", "100%", "[%]", "[Status]", ""),
        ("Rollback Tested (12 mo)", "100%", "[%]", "[Status]", ""),
        ("Reviews Overdue", "0", "[#]", "[Status]", ""),
    ]

    row = 42
    for metric in mitigation_metrics:
        for col, value in enumerate(metric, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Evidence Protection Summary
    ws.cell(row=48, column=1, value="AUDIT EVIDENCE PROTECTION").font = Font(bold=True, size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=49, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    evidence_metrics = [
        ("Evidence Categories", "-", "[#]", "-", ""),
        ("Encryption at Rest", "100%", "[%]", "[Status]", ""),
        ("Encryption in Transit", "100%", "[%]", "[Status]", ""),
        ("Access Reviews Current", "100%", "[%]", "[Status]", ""),
    ]

    row = 50
    for metric in evidence_metrics:
        for col, value in enumerate(metric, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Column widths
    widths = [35, 15, 15, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_gap_analysis_sheet(ws):
    """Create the Gap Analysis sheet."""
    ws.title = "Gap_Analysis"

    ws.merge_cells('A1:J1')
    title_cell = ws.cell(row=1, column=1, value="CONSOLIDATED GAP ANALYSIS")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Gaps identified across all Testing and Audit Protection assessments")

    # Test Data Protection Gaps
    ws.cell(row=4, column=1, value="TEST DATA PROTECTION GAPS (IMP-A.8.33-34.1)").font = Font(bold=True, size=12)

    headers = ["Gap_ID", "Gap_Description", "Area", "Severity", "Current_State", "Target_State", "Owner", "Target_Date", "Status", "Evidence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(severity_dv)
    severity_dv.add('D6:D55')
    severity_dv.add('D58:D107')

    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Closed"')
    ws.add_data_validation(status_dv)
    status_dv.add('I6:I55')
    status_dv.add('I58:I107')

    # Format input rows for test data gaps
    for row in range(6, 56):
        ws.cell(row=row, column=1, value=f"GAP-TD-{str(row-5).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Audit Activity Management Gaps
    ws.cell(row=57, column=1, value="AUDIT ACTIVITY MANAGEMENT GAPS (IMP-A.8.33-34.2)").font = Font(bold=True, size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=58, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Format input rows for audit gaps
    for row in range(59, 109):
        ws.cell(row=row, column=1, value=f"GAP-AA-{str(row-58).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Gap Statistics
    ws.cell(row=111, column=1, value="GAP ANALYSIS STATISTICS").font = Font(bold=True, size=12)

    stats_headers = ["Metric", "Count", "Percentage"]
    for col, header in enumerate(stats_headers, 1):
        cell = ws.cell(row=112, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    stats = [
        ("Total Gaps Identified", "[#]", "100%"),
        ("Critical Priority", "[#]", "[%]"),
        ("High Priority", "[#]", "[%]"),
        ("Medium Priority", "[#]", "[%]"),
        ("Low Priority", "[#]", "[%]"),
        ("Status - Open", "[#]", "[%]"),
        ("Status - In Progress", "[#]", "[%]"),
        ("Status - Closed", "[#]", "[%]"),
        ("Overdue Gaps", "[#]", "[%]"),
    ]

    row = 113
    for stat in stats:
        for col, value in enumerate(stat, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Column widths
    widths = [15, 50, 20, 12, 30, 30, 20, 15, 15, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_exception_register_sheet(ws):
    """Create the Exception Register sheet."""
    ws.title = "Exception_Register"

    ws.merge_cells('A1:T1')
    title_cell = ws.cell(row=1, column=1, value="EXCEPTION REGISTER")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Approved exceptions to testing and audit protection requirements")

    headers = [
        "Exception_ID", "Exception_Title", "Policy_Requirement", "Assessment_Source",
        "Business_Justification", "Risk_Assessment", "Compensating_Controls",
        "Requested_By", "Request_Date", "Approval_Status", "Approved_By",
        "Approval_Date", "Expiration_Date", "Renewal_Required", "Risk_Level",
        "Risk_Owner", "Review_Date", "Next_Review", "Evidence_Reference", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    source_dv = DataValidation(type="list", formula1='"IMP-A.8.33-34.1,IMP-A.8.33-34.2"')
    ws.add_data_validation(source_dv)
    source_dv.add('D5:D34')

    status_dv = DataValidation(type="list", formula1='"Approved,Pending,Denied,Expired"')
    ws.add_data_validation(status_dv)
    status_dv.add('J5:J34')

    yn_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('N5:N34')

    risk_dv = DataValidation(type="list", formula1='"High,Medium,Low"')
    ws.add_data_validation(risk_dv)
    risk_dv.add('O5:O34')

    # Format input rows
    for row in range(5, 35):
        ws.cell(row=row, column=1, value=f"EXC-{str(row-4).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics
    ws.cell(row=37, column=1, value="EXCEPTION STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Exceptions", "=COUNTA(B5:B34)"),
        ("Approved Exceptions", '=COUNTIF(J5:J34,"Approved")'),
        ("Pending Exceptions", '=COUNTIF(J5:J34,"Pending")'),
        ("Expired Exceptions", '=COUNTIF(J5:J34,"Expired")'),
        ("High Risk Exceptions", '=COUNTIF(O5:O34,"High")'),
    ]

    row = 39
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [12, 35, 35, 25, 50, 40, 50, 25, 12, 18, 25, 12, 12, 12, 15, 25, 12, 12, 20, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_kpis_metrics_sheet(ws):
    """Create the KPIs and Metrics sheet."""
    ws.title = "KPIs_and_Metrics"

    ws.merge_cells('A1:I1')
    title_cell = ws.cell(row=1, column=1, value="KEY PERFORMANCE INDICATORS & METRICS")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Detailed tracking with historical trends")

    # Test Data KPIs
    ws.cell(row=4, column=1, value="TEST DATA PROTECTION KPIs").font = Font(bold=True, size=12)

    headers = ["KPI", "Target", "Q1", "Q2", "Q3", "Q4", "Current", "Status", "Trend"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    test_data_kpis = [
        ("Data Set Authorization Rate", "100%"),
        ("PII Masking Coverage", "100%"),
        ("Average Masking Effectiveness", ">=4.0"),
        ("Environment Security Compliance", ">=90%"),
        ("Refresh Authorization Rate", "100%"),
        ("High Re-identification Risk Count", "0"),
    ]

    row = 6
    for kpi, target in test_data_kpis:
        ws.cell(row=row, column=1, value=kpi).border = THIN_BORDER
        ws.cell(row=row, column=2, value=target).border = THIN_BORDER
        for col in range(3, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        row += 1

    # Status validation
    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant"')
    ws.add_data_validation(status_dv)
    status_dv.add('H6:H30')

    # Trend validation
    trend_dv = DataValidation(type="list", formula1='"Up,Stable,Down"')
    ws.add_data_validation(trend_dv)
    trend_dv.add('I6:I30')

    # Audit Activity KPIs
    ws.cell(row=14, column=1, value="AUDIT ACTIVITY MANAGEMENT KPIs").font = Font(bold=True, size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=15, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    audit_kpis = [
        ("Audit Authorization Rate", "100%"),
        ("Tool Authorization Rate", "100%"),
        ("Access Approval Rate", "100%"),
        ("Access Revocation Timeliness", "100%"),
        ("NDA Coverage", "100%"),
        ("Access Logging Coverage", "100%"),
        ("MFA Coverage", "100%"),
        ("Critical System Mitigation Coverage", "100%"),
        ("Evidence Encryption Rate", "100%"),
        ("Audit Follow-up Closure Rate", ">=90%"),
    ]

    row = 16
    for kpi, target in audit_kpis:
        ws.cell(row=row, column=1, value=kpi).border = THIN_BORDER
        ws.cell(row=row, column=2, value=target).border = THIN_BORDER
        for col in range(3, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        row += 1

    # Combined KPIs
    ws.cell(row=28, column=1, value="COMBINED COMPLIANCE KPIs").font = Font(bold=True, size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=29, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    combined_kpis = [
        ("Overall Compliance Rate", ">=85%"),
        ("Critical Gaps Open", "0"),
        ("High Gaps Open", "0"),
        ("Evidence Coverage", "100%"),
        ("Exception Count", "Minimize"),
        ("Audit Readiness Score", ">=85%"),
    ]

    row = 30
    for kpi, target in combined_kpis:
        ws.cell(row=row, column=1, value=kpi).border = THIN_BORDER
        ws.cell(row=row, column=2, value=target).border = THIN_BORDER
        for col in range(3, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        row += 1

    # Column widths
    widths = [40, 12, 12, 12, 12, 12, 12, 15, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_remediation_roadmap_sheet(ws):
    """Create the Remediation Roadmap sheet."""
    ws.title = "Remediation_Roadmap"

    ws.merge_cells('A1:K1')
    title_cell = ws.cell(row=1, column=1, value="REMEDIATION ROADMAP")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Structured plan to address identified gaps")

    # Timeline Overview
    ws.cell(row=4, column=1, value="PHASE TIMELINE").font = Font(bold=True, size=12)

    phase_headers = ["Phase", "Timeline", "Focus Areas", "Expected Outcome"]
    for col, header in enumerate(phase_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    phases = [
        ("Phase 1 - Critical (0-30 days)", "[Dates]", "Critical gaps", "All critical gaps closed"),
        ("Phase 2 - High Priority (31-90 days)", "[Dates]", "High priority gaps", "Major risks mitigated"),
        ("Phase 3 - Medium Priority (91-180 days)", "[Dates]", "Medium priority gaps", "Full compliance achieved"),
        ("Phase 4 - Continuous Improvement (Ongoing)", "[Dates]", "Low priority and enhancements", "Best practice adoption"),
    ]

    row = 6
    for phase in phases:
        for col, value in enumerate(phase, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Remediation Actions
    ws.cell(row=12, column=1, value="REMEDIATION ACTIONS").font = Font(bold=True, size=12)

    action_headers = ["Action_ID", "Description", "Priority", "Source", "Owner", "Start", "Target", "Status", "Progress", "Dependencies", "Notes"]
    for col, header in enumerate(action_headers, 1):
        cell = ws.cell(row=13, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    priority_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(priority_dv)
    priority_dv.add('C14:C73')

    source_dv = DataValidation(type="list", formula1='"IMP-A.8.33-34.1,IMP-A.8.33-34.2,Both"')
    ws.add_data_validation(source_dv)
    source_dv.add('D14:D73')

    status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,Completed,Blocked"')
    ws.add_data_validation(status_dv)
    status_dv.add('H14:H73')

    # Format input rows
    for row in range(14, 74):
        ws.cell(row=row, column=1, value=f"RA-{str(row-13).zfill(3)}")
        for col in range(1, len(action_headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Progress Summary
    ws.cell(row=76, column=1, value="PROGRESS SUMMARY").font = Font(bold=True, size=12)

    progress_headers = ["Metric", "Count", "Percentage", "Target", "Status"]
    for col, header in enumerate(progress_headers, 1):
        cell = ws.cell(row=77, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    progress_items = [
        ("Total Actions", "[#]", "100%", "N/A", "-"),
        ("Completed", "[#]", "[%]", "100%", "[Status]"),
        ("In Progress", "[#]", "[%]", "-", "-"),
        ("Not Started", "[#]", "[%]", "-", "[Status]"),
        ("Overdue", "[#]", "[%]", "<10%", "[Status]"),
        ("Critical Remaining", "[#]", "[%]", "0%", "[Status]"),
        ("Average Progress", "[%]", "N/A", ">=70%", "[Status]"),
    ]

    row = 78
    for item in progress_items:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if value.startswith("["):
                cell.fill = INPUT_FILL
        row += 1

    # Column widths
    widths = [12, 50, 12, 20, 20, 12, 12, 15, 12, 30, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_evidence_consolidation_sheet(ws):
    """Create the Evidence Consolidation sheet."""
    ws.title = "Evidence_Consolidation"

    ws.merge_cells('A1:I1')
    title_cell = ws.cell(row=1, column=1, value="CONSOLIDATED EVIDENCE REGISTER")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Evidence from both Testing and Audit Protection assessments (200 entries)")

    headers = [
        "Evidence_ID", "Source_Assessment", "Evidence_Type", "Description",
        "Location", "Date_Collected", "Collected_By", "Verification_Status", "Auditor_Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    source_dv = DataValidation(type="list", formula1='"IMP-A.8.33-34.1,IMP-A.8.33-34.2"')
    ws.add_data_validation(source_dv)
    source_dv.add('B5:B204')

    type_dv = DataValidation(type="list", formula1='"Policy Document,Technical Config,Authorization Record,Audit Log,Verification Report,Training Record,Other"')
    ws.add_data_validation(type_dv)
    type_dv.add('C5:C204')

    status_dv = DataValidation(type="list", formula1='"Verified,Pending,Not Verified"')
    ws.add_data_validation(status_dv)
    status_dv.add('H5:H204')

    # Format input rows
    for row in range(5, 205):
        ws.cell(row=row, column=1, value=f"EV-{str(row-4).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics
    ws.cell(row=207, column=1, value="EVIDENCE STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Evidence", "=COUNTA(B5:B204)"),
        ("From Test Data Assessment", '=COUNTIF(B5:B204,"IMP-A.8.33-34.1")'),
        ("From Audit Assessment", '=COUNTIF(B5:B204,"IMP-A.8.33-34.2")'),
        ("Verified", '=COUNTIF(H5:H204,"Verified")'),
        ("Pending Verification", '=COUNTIF(H5:H204,"Pending")'),
    ]

    row = 209
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [15, 20, 25, 50, 40, 15, 25, 18, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_ciso_certification_sheet(ws):
    """Create the CISO Certification sheet."""
    ws.title = "CISO_Certification"

    ws.merge_cells('A1:D1')
    title_cell = ws.cell(row=1, column=1, value="CISO COMPLIANCE CERTIFICATION")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Formal attestation of Testing and Audit Protection compliance status")

    # Certification Summary
    ws.cell(row=4, column=1, value="CERTIFICATION SUMMARY").font = Font(bold=True, size=12)

    summary_items = [
        ("Dashboard Document:", "ISMS-IMP-A.8.33-34.3 - Compliance Dashboard"),
        ("Certification Period:", "[USER INPUT]"),
        ("Overall Compliance Rate:", "[Formula from Executive_Dashboard]"),
        ("Critical Gaps:", "[Formula]"),
        ("High Gaps:", "[Formula]"),
        ("Active Exceptions:", "[Formula]"),
        ("Assessment 1 Status:", "[Formula from IMP-A.8.33-34.1]"),
        ("Assessment 2 Status:", "[Formula from IMP-A.8.33-34.2]"),
        ("Total Evidence:", "[Formula]"),
        ("Audit Readiness:", "[Formula based on compliance]"),
        ("Certification Status:", "[Dropdown]"),
    ]

    row = 6
    for label, value in summary_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value or "[Formula" in value or "[Dropdown]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Certification status dropdown
    cert_status_dv = DataValidation(type="list", formula1='"Certified,Certified with Conditions,Not Certified"')
    ws.add_data_validation(cert_status_dv)
    cert_status_dv.add('B16')

    # CISO Attestation
    ws.cell(row=19, column=1, value="CISO ATTESTATION").font = Font(bold=True, size=12)

    ws.merge_cells('A21:D21')
    ws.cell(row=21, column=1, value="I, _________________________________ (CISO Name), certify that:")
    ws.cell(row=21, column=1).font = Font(italic=True)

    attestation_items = [
        "1. Assessment Accuracy",
        "   [ ] I have reviewed the source assessments (IMP-A.8.33-34.1 and IMP-A.8.33-34.2)",
        "   [ ] I confirm the data in this dashboard accurately reflects our compliance status",
        "   [ ] Assessment methodologies are appropriate and consistently applied",
        "",
        "2. Gap Understanding",
        "   [ ] I understand all identified gaps and their business impact",
        "   [ ] Critical gaps have approved remediation plans with adequate resources",
        "   [ ] High-priority gaps are being actively addressed",
        "",
        "3. Exception Acceptance",
        "   [ ] I have reviewed all active exceptions",
        "   [ ] I accept the residual risk associated with each exception",
        "   [ ] Compensating controls are adequate and monitored",
        "",
        "4. Risk Acceptance",
        "   [ ] I accept the residual risk documented in this assessment",
        "   [ ] Risk levels are within organizational risk appetite",
        "   [ ] Additional risk treatment is not required at this time",
        "",
        "5. Audit Readiness",
        "   [ ] The organization is/is not ready for ISO 27001 audit of Controls A.8.33-34",
        "   [ ] Evidence is complete and accessible",
        "   [ ] Personnel are prepared for audit interviews",
    ]

    row = 23
    for item in attestation_items:
        ws.cell(row=row, column=1, value=item)
        if item.startswith("   [ ]"):
            ws.cell(row=row, column=1).font = Font(size=10)
        elif item and not item.startswith("   "):
            ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

    # Certification Decision
    row += 1
    ws.cell(row=row, column=1, value="Certification Decision:").font = Font(bold=True)
    cell = ws.cell(row=row, column=2, value="[Dropdown]")
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    cert_status_dv.add(f'B{row}')

    row += 2
    ws.cell(row=row, column=1, value="Conditions (if applicable):").font = Font(bold=True)
    row += 1
    ws.merge_cells(f'A{row}:D{row+2}')
    cell = ws.cell(row=row, column=1, value="[Enter conditions here]")
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER

    row += 4
    ws.cell(row=row, column=1, value="CISO Signature:").font = Font(bold=True)
    cell = ws.cell(row=row, column=2, value="_________________________________")
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER

    row += 1
    ws.cell(row=row, column=1, value="Date:").font = Font(bold=True)
    cell = ws.cell(row=row, column=2, value="[DD.MM.YYYY]")
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER

    row += 2
    ws.cell(row=row, column=1, value="Certification Valid Until:").font = Font(bold=True)
    cell = ws.cell(row=row, column=2, value="[Formula: Date + 90 days]")
    cell.fill = LOCKED_FILL
    cell.border = THIN_BORDER

    row += 1
    ws.cell(row=row, column=1, value="Next Certification Due:").font = Font(bold=True)
    cell = ws.cell(row=row, column=2, value="[Formula: Date + 90 days]")
    cell.fill = LOCKED_FILL
    cell.border = THIN_BORDER

    # Certification History
    row += 3
    ws.cell(row=row, column=1, value="CERTIFICATION HISTORY").font = Font(bold=True, size=12)

    history_headers = ["Certification_ID", "Date", "CISO", "Status", "Overall_Compliance", "Critical_Gaps", "Conditions", "Valid_Until"]
    row += 1
    for col, header in enumerate(history_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Format history rows
    for r in range(row + 1, row + 11):
        ws.cell(row=r, column=1, value=f"CERT-{str(r-row).zfill(3)}")
        for col in range(1, len(history_headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


def create_approval_sign_off_sheet(ws):
    """Create the Approval Sign-Off sheet."""
    ws.title = "Approval_Sign_Off"

    ws.merge_cells('A1:D1')
    title_cell = ws.cell(row=1, column=1, value="DASHBOARD APPROVAL & SIGN-OFF")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Executive approval workflow for compliance dashboard")

    # Dashboard Summary
    ws.cell(row=4, column=1, value="DASHBOARD SUMMARY").font = Font(bold=True, size=12)

    summary_items = [
        ("Document ID:", DOCUMENT_ID),
        ("Report Type:", "Compliance Summary Dashboard"),
        ("Reporting Period:", "[USER INPUT]"),
        ("Overall Compliance:", "[Formula]"),
        ("Test Data Compliance:", "[Formula]"),
        ("Audit Activity Compliance:", "[Formula]"),
        ("Critical Findings:", "[Formula]"),
        ("Active Exceptions:", "[Formula]"),
        ("Evidence Entries:", "[Formula]"),
        ("Dashboard Status:", "[Dropdown]"),
    ]

    row = 6
    for label, value in summary_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value or "[Dropdown]" in value or "[Formula]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Status dropdown
    status_dv = DataValidation(type="list", formula1='"Final,Requires Action,Draft"')
    ws.add_data_validation(status_dv)
    status_dv.add('B15')

    # Prepared By
    row += 2
    ws.cell(row=row, column=1, value="PREPARED BY").font = Font(bold=True, size=12)
    row += 2

    prepared_items = [
        ("Name:", "[USER INPUT]"),
        ("Role/Title:", "[USER INPUT]"),
        ("Department:", "[USER INPUT]"),
        ("Email:", "[USER INPUT]"),
        ("Date Prepared:", "[DD.MM.YYYY]"),
        ("Signature:", "[USER INPUT]"),
    ]

    for label, value in prepared_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Reviewed By
    row += 2
    ws.cell(row=row, column=1, value="REVIEWED BY").font = Font(bold=True, size=12)
    row += 2

    reviewed_items = [
        ("Name:", "[USER INPUT]"),
        ("Role/Title:", "Security Manager / Compliance Officer"),
        ("Review Date:", "[DD.MM.YYYY]"),
        ("Review Notes:", "[Text area]"),
        ("Data Accuracy:", "[Dropdown]"),
        ("Recommendation:", "[Dropdown]"),
        ("Conditions:", "[Text area]"),
        ("Signature:", "[USER INPUT]"),
    ]

    for label, value in reviewed_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value or "[DD.MM.YYYY]" in value or "[Text area]" in value or "[Dropdown]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Dropdowns for reviewed by
    acc_dv = DataValidation(type="list", formula1='"Confirmed,Minor Issues,Major Issues"')
    ws.add_data_validation(acc_dv)
    acc_dv.add('B36')

    rec_dv = DataValidation(type="list", formula1='"Approve,Approve with Conditions,Reject"')
    ws.add_data_validation(rec_dv)
    rec_dv.add('B37')

    # Approved By - Executive
    row += 2
    ws.cell(row=row, column=1, value="APPROVED BY - EXECUTIVE").font = Font(bold=True, size=12)
    row += 2

    approved_items = [
        ("Name:", "[USER INPUT]"),
        ("Role/Title:", "CISO / CRO / CEO"),
        ("Approval Date:", "[DD.MM.YYYY]"),
        ("Approval Decision:", "[Dropdown]"),
        ("Conditions/Notes:", "[Text area]"),
        ("Signature:", "[USER INPUT]"),
    ]

    for label, value in approved_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value or "[DD.MM.YYYY]" in value or "[Text area]" in value or "[Dropdown]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Approval decision dropdown
    app_dec_dv = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected"')
    ws.add_data_validation(app_dec_dv)
    app_dec_dv.add('B47')

    # Next Review
    row += 2
    ws.cell(row=row, column=1, value="NEXT REVIEW").font = Font(bold=True, size=12)
    row += 2

    next_items = [
        ("Next Review Date:", "[Formula: Approval Date + 90 days]"),
        ("Review Responsibility:", "[USER INPUT]"),
        ("Focus Areas:", "[USER INPUT]"),
        ("Distribution List:", "[USER INPUT]"),
    ]

    for label, value in next_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30


def generate_workbook():
    """Generate the complete assessment workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Create all sheets
    create_executive_dashboard_sheet(wb.create_sheet())
    create_test_data_compliance_sheet(wb.create_sheet())
    create_audit_activity_compliance_sheet(wb.create_sheet())
    create_gap_analysis_sheet(wb.create_sheet())
    create_exception_register_sheet(wb.create_sheet())
    create_kpis_metrics_sheet(wb.create_sheet())
    create_remediation_roadmap_sheet(wb.create_sheet())
    create_evidence_consolidation_sheet(wb.create_sheet())
    create_ciso_certification_sheet(wb.create_sheet())
    create_approval_sign_off_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - INITIAL CREATION
# QA_TOOL: Claude Code
# CHANGES: Initial generator creation for A.8.33-34 Testing and Audit Protection
# =============================================================================
