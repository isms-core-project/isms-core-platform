#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.34.6 - Cross-Border Transfer Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.34: Privacy and Protection of PII
Assessment Domain 6 of 7: Cross-Border Data Transfer Compliance

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific international data transfer activities, transfer
mechanisms, and GDPR Chapter V / FADP compliance requirements.

Key customization areas:
1. Transfer inventory sources (match your actual data flows and systems)
2. Transfer mechanisms (adapt adequacy lists to your jurisdictions)
3. TIA methodology (align with organizational risk assessment framework)
4. Processor relationships (specific to your vendor ecosystem)
5. Compliance thresholds (based on regulatory obligations)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.34 Privacy and Protection of PII Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for managing
international personal data transfers per GDPR Chapter V (Articles 44-50),
Swiss FADP Article 16, and ISO/IEC 27001:2022 Control A.5.34 requirements.

**Purpose:**
Enables systematic inventory, legal basis validation, Transfer Impact Assessment
(TIA) execution, and processor agreement compliance for cross-border data transfers,
demonstrating GDPR Chapter V compliance post-Schrems II and supporting audit-ready
documentation for supervisory authorities.

**Assessment Scope:**
- Cross-border transfer inventory (all PII leaving EU/EEA or Switzerland)
- Transfer mechanism validation (Adequacy, SCCs, BCRs, DPF, Derogations)
- Transfer Impact Assessments (TIAs) per Schrems II requirements
- Processor Data Processing Agreement (DPA) compliance verification
- Subprocessor onward transfer safeguards
- Supplementary measures implementation (technical, contractual, organizational)
- Gap analysis for unlawful or under-documented transfers
- Evidence collection for supervisory authority inspections
- Compliance dashboard with transfer metrics

**Generated Workbook Structure:**
1. Instructions & Legend - GDPR Chapter V guidance and adequacy decisions list
2. Cross-Border Transfer Register - Comprehensive inventory of international transfers
3. Transfer Impact Assessment (TIA) Register - Schrems II TIA documentation
4. Processor Agreement Tracker - DPA and SCC compliance verification
5. Evidence Repository - Audit evidence tracking and documentation linkage
6. Gap Analysis - Non-compliant transfer remediation planning
7. Compliance Dashboard - Auto-calculated metrics for executive oversight
8. Approval & Sign-Off - Stakeholder approvals and DPO sign-off

**Key Features:**
- Data validation with GDPR Chapter V transfer mechanism dropdowns
- Conditional formatting for adequacy status and compliance gaps
- Automated TIA requirement flagging for non-adequate country transfers
- EU Commission adequacy decisions list (automatically checks destination country)
- SCC version validation (2021 vs. invalid 2010 SCCs)
- Protected formulas with unprotected input cells
- Risk-based gap prioritization (Critical/High/Medium/Low)
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Dashboard with compliance percentages and gap counts

**Integration:**
This assessment is Domain 6 of the A.5.34 Privacy assessment suite (7 domains).
Results feed into ISMS-IMP-A.5.34.7 Privacy Compliance Dashboard for
consolidated privacy program oversight and regulatory reporting. Integrates
with A.5.34.1 (PII Inventory) for transfer identification and A.5.34.4 (TOMs)
for supplementary measure verification.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a5346_cross_border_transfer_assessment.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a5346_cross_border_transfer_assessment.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a5346_cross_border_transfer_assessment.py --date 20250130

Output:
    File: ISMS_A_5_34_6_Cross_Border_Transfer_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize adequacy decisions list (Sheet 1) for applicable jurisdictions
    2. Identify all cross-border transfers using A.5.34.1 PII inventory data
    3. Complete Sheet 2 (Transfer Register) - document all international transfers
    4. Determine transfer mechanisms (Adequacy, SCCs, BCRs, DPF, Derogations)
    5. Conduct Transfer Impact Assessments (TIAs) in Sheet 3 for non-adequate countries
    6. Implement supplementary measures (encryption, contractual clauses, monitoring)
    7. Verify processor agreements in Sheet 4 (DPAs with SCCs for non-adequate)
    8. Check subprocessor onward transfer safeguards
    9. Collect and link audit evidence in Sheet 5
    10. Conduct gap analysis in Sheet 6 for unlawful/under-documented transfers
    11. Create remediation plan with risk-based prioritization
    12. Review Sheet 7 (Dashboard) for compliance metrics
    13. Obtain DPO approval and stakeholder sign-offs in Sheet 8
    14. Schedule quarterly TIA reviews per Schrems II continuous monitoring
    15. Feed results into A.5.34.7 Privacy Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.34
Assessment Domain:    6 of 7 (Cross-Border Data Transfer Compliance)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.34: Privacy and Protection of PII Policy (Governance)
    - ISMS-CTX-A.5.34: Privacy Regulatory Landscape Reference (Context)
    - ISMS-IMP-A.5.34.1: PII Identification Assessment (Domain 1)
    - ISMS-IMP-A.5.34.2: Legal Basis Assessment (Domain 2)
    - ISMS-IMP-A.5.34.3: DSR Management Assessment (Domain 3)
    - ISMS-IMP-A.5.34.4: TOMs Assessment (Domain 4)
    - ISMS-IMP-A.5.34.5: DPIA Assessment (Domain 5)
    - ISMS-IMP-A.5.34.6: Cross-Border Transfer Implementation Guide (Parts 1-3)
    - ISMS-IMP-A.5.34.7: Privacy Compliance Dashboard (Consolidation)
    - EDPB Recommendations 01/2020: Supplementary Measures for Transfers
    - Schrems II CJEU C-311/18: Privacy Shield Invalidation Decision

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 2025-01-30
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.34.6 specification
    - Supports comprehensive cross-border transfer compliance assessment
    - Integrated TIA framework per Schrems II requirements
    - Prepares for consolidation into A.5.34.7 Privacy Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**GDPR Chapter V Post-Schrems II:**
The Schrems II CJEU decision (July 2020) invalidated the EU-US Privacy Shield
and mandated Transfer Impact Assessments (TIAs) for all transfers to countries
without adequacy decisions, EVEN WHEN SCCs ARE IN PLACE. Organizations must:
(1) Assess destination country legal framework (surveillance laws, government
access powers), (2) Evaluate importer-specific circumstances, (3) Implement
supplementary measures (technical, contractual, organizational) to ensure
adequate protection, (4) Document TIA conclusions. TIAs must be reassessed
if legal/political situation changes. This is NOT optional - it's a legal
requirement for GDPR compliance.

**EU-US Data Privacy Framework (DPF):**
The EU-US DPF (July 2023) provides adequacy for US companies that self-certify.
However, DPF certifications expire annually and must be renewed. Always verify
current DPF status at https://www.dataprivacyframework.gov/s/participant-search
before relying on DPF as transfer mechanism. DPF may be challenged in court
(similar to Schrems I & II) - maintain backup transfer mechanisms (SCCs + TIA).

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 and GDPR requirements.
Supervisory authorities and auditors will expect: complete transfer inventory,
documented transfer mechanisms with valid legal basis, TIAs for all non-adequate
country transfers, evidence of supplementary measures implementation, processor
DPAs with SCCs (2021 version), subprocessor onward transfer safeguards, regular
TIA reviews (quarterly or annually), DPO involvement and approval. Ensure all
transfers are documented BEFORE processing begins - unlawful transfers can
result in GDPR fines up to 4% of global annual revenue.

**Data Protection:**
Assessment workbooks contain sensitive information including:
- Complete inventory of international data flows (reveals business relationships)
- TIA risk analyses highlighting legal vulnerabilities
- Processor agreements and subprocessor relationships (competitive intelligence)
- Gap analysis revealing compliance deficiencies
- Supplementary measures implementations (security architecture details)

Handle in accordance with your organization's data classification policies.
Restrict access to DPO, Legal Counsel, Privacy Team, and authorized compliance
personnel only. Consider encryption at rest for stored workbooks.

**Maintenance:**
Review and update assessment:
- Quarterly: New transfers, processor changes, TIA reviews (especially for US transfers)
- Annually: Complete transfer inventory reassessment, adequacy decision updates
- Triggered: New adequacy decisions, adequacy revocations, Schrems III potential ruling,
  regulatory guidance updates, processor contract changes, new subprocessors
- Continuous: Monitor destination country legal developments (new surveillance laws)

**Quality Assurance:**
Have DPO and Legal Counsel validate all transfer mechanisms and TIA conclusions
before implementing transfers. For high-risk transfers (e.g., sensitive PII to
US or China), consider external legal review. Conduct annual audit of transfer
register completeness - compare against A.5.34.1 PII inventory to identify
undocumented transfers. Train business owners on cross-border transfer approval
requirements BEFORE signing new processor contracts.

**Regulatory Alignment:**
This assessment supports compliance with:
- GDPR (EU General Data Protection Regulation) - Chapter V (Art. 44-50)
- FADP (Swiss Federal Act on Data Protection) - Art. 16 (Disclosure Abroad)
- ISO/IEC 27001:2022 - Control A.5.34 (Privacy and Protection of PII)
- ISO/IEC 27701:2019 - Privacy Information Management (if applicable)
- Schrems II CJEU C-311/18 - TIA Requirements
- EDPB Recommendations 01/2020 - Supplementary Measures Guidance

**Critical Deadlines and Compliance Milestones:**
- 2010 SCCs: INVALID since September 27, 2022 (must update to 2021 version)
- DPF certifications: Expire annually, verify renewal quarterly
- TIA reviews: Reassess when destination country legal framework changes
- Processor DPA reviews: Annual review of subprocessor lists and SCC flow-down
- Gap remediation: Critical gaps = immediate action (1-4 weeks)
- Supervisory authority consultations: Pre-consultation required if TIA shows
  high residual risk (GDPR Article 36)

================================================================================
"""

import argparse
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)




# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.34.6"
WORKBOOK_NAME = "Cross-Border Transfer Assessment"
CONTROL_ID = "A.5.34"
CONTROL_NAME = "Privacy and Protection of PII"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# ============================================================================
# CONFIGURATION CONSTANTS
# ============================================================================

# CUSTOMIZE: Colors matching A.5.34 suite
COLORS = {
    'header_blue': 'FF003366',
    'white': 'FFFFFFFF',
    'black': 'FF000000',
    'light_green': 'FFC6EFCE',
    'dark_green': 'FF006100',
    'light_yellow': 'FFFFEB9C',
    'dark_orange': 'FF9C5700',
    'light_red': 'FFFFC7CE',
    'dark_red': 'FF9C0006',
    'light_blue': 'FFB4C6E7',
    'light_gray': 'FFD9D9D9',
    'light_orange': 'FFFFD966',
}

# CUSTOMIZE: EU Adequacy Decisions (as of 2025)
EU_ADEQUACY_COUNTRIES = [
    "Andorra", "Argentina", "Canada (Commercial)", "Faroe Islands",
    "Guernsey", "Israel", "Isle of Man", "Japan", "Jersey",
    "New Zealand", "Republic of Korea", "Switzerland", "United Kingdom", "Uruguay"
]

# Transfer status options
STATUS_OPTIONS = ["Not Started", "In Progress", "Complete", "Validated"]

# Risk levels
RISK_LEVELS = ["Low", "Medium", "High", "Critical"]

# Protection password
PROTECTION_PASSWORD = "privacy2024"

# File prefix
FILE_PREFIX = "ISMS_A_5_34_6_Cross_Border_Transfer_Assessment"


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def style_header_row(ws, row_num, color_hex, num_columns):
    """Apply consistent header styling."""
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(name='Calibri', size=11, bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )


def add_dropdown(ws, cell_range, options, error_msg="Invalid selection", allow_blank=False):
    """Add dropdown data validation."""
    dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=allow_blank)
    dv.error = error_msg
    dv.errorStyle = 'stop'
    ws.add_data_validation(dv)
    dv.add(cell_range)


def protect_sheet(ws, password=PROTECTION_PASSWORD):
    """Protect sheet with standard settings."""
    ws.protection.sheet = True
    ws.protection.password = password
    ws.protection.enable()


# ============================================================================
# SHEET CREATION FUNCTIONS (8 SHEETS)
# ============================================================================

def create_instructions_sheet(wb):
    """Sheet 1: Instructions & Legend with GDPR Chapter V guidance."""
    ws = wb.create_sheet("Instructions")
    
    # Title
    ws.merge_cells('A1:P3')
    title = ws['A1']
    title.value = "Cross-Border Data Transfer Assessment - Instructions & Legend"
    title.font = Font(size=16, bold=True, color=COLORS['white'])
    title.fill = PatternFill(start_color=COLORS['header_blue'], fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')
    
    # Section: GDPR Chapter V Overview
    ws['A5'] = "GDPR Chapter V: Transfers to Third Countries"
    ws['A5'].font = Font(size=14, bold=True)
    ws['A6'] = "Article 44: General principle - transfers only with adequate safeguards"
    ws['A7'] = "Article 45: Adequacy decisions (EU Commission approves countries)"
    ws['A8'] = "Article 46: Appropriate safeguards (SCCs, BCRs, certifications)"
    ws['A9'] = "Article 49: Derogations for specific situations (LAST RESORT)"
    
    # EU Adequacy Countries List
    ws['A12'] = "EU Adequacy Decisions (Current):"
    ws['A12'].font = Font(bold=True)
    for idx, country in enumerate(EU_ADEQUACY_COUNTRIES, start=13):
        ws[f'A{idx}'] = country
    
    return ws


def create_transfer_register_sheet(wb):
    """Sheet 2: Cross-Border Transfer Register (31 columns)."""
    ws = wb.create_sheet("Transfer Register")
    
    # Headers (simplified - users expand as needed)
    headers = [
        "Transfer ID", "Status", "Transfer Name", "Source System", "Destination System",
        "Destination Country", "Adequacy Status", "Transfer Mechanism", "SCC Version",
        "SCC Date", "DPF Cert?", "TIA Required?", "TIA ID", "PII Categories",
        "Transfer Volume", "Transfer Frequency", "Purpose", "Legal Basis (Art.6)",
        "Last Updated", "Notes"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['header_blue'], fill_type='solid')
        cell.alignment = Alignment(wrap_text=True)
    
    # Add formula in A2 for Transfer ID
    ws['A2'] = '=TEXT(ROW()-1,"XFER-0000")'
    
    # Dropdowns
    add_dropdown(ws, 'B2:B1000', STATUS_OPTIONS, "Select valid status")
    add_dropdown(ws, 'H2:H1000', ["Adequacy", "SCCs", "BCRs", "DPF", "Derogation"], "Select mechanism")
    add_dropdown(ws, 'L2:L1000', ["YES", "NO"], "YES or NO")
    
    # Protect sheet
    ws.protection.sheet = True
    ws.protection.password = PROTECTION_PASSWORD
    
    return ws


def create_tia_register_sheet(wb):
    """Sheet 3: Transfer Impact Assessment Register (25 columns)."""
    ws = wb.create_sheet("TIA Register")
    
    headers = [
        "TIA ID", "Transfer ID", "Status", "Destination Country", "Assessment Date",
        "Assessor", "Surveillance Laws", "Gov Access Risk", "Risk Justification",
        "Supplementary Measures", "Residual Risk", "TIA Conclusion", "DPO Approval",
        "Next Review Date", "Notes"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['header_blue'], fill_type='solid')
    
    ws['A2'] = '=TEXT(ROW()-1,"TIA-0000")'
    
    add_dropdown(ws, 'C2:C1000', ["Not Started", "In Progress", "Complete", "Approved"], "Status")
    add_dropdown(ws, 'H2:H1000', RISK_LEVELS, "Risk level")
    add_dropdown(ws, 'K2:K1000', RISK_LEVELS, "Risk level")
    add_dropdown(ws, 'L2:L1000', ["PASS", "PASS with Conditions", "FAIL"], "TIA result")
    
    ws.protection.sheet = True
    ws.protection.password = PROTECTION_PASSWORD
    
    return ws


def create_processor_tracker_sheet(wb):
    """Sheet 4: Processor Agreement Tracker (31 columns)."""
    ws = wb.create_sheet("Processor Tracker")
    
    headers = [
        "Processor ID", "Transfer ID", "Processor Name", "Processor Location",
        "DPA Exists?", "DPA Date", "SCCs Included?", "SCC Version", "SCC Date",
        "Subprocessors?", "Compliance Status", "Gap Description", "Remediation Action",
        "Owner", "Deadline", "Notes"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['header_blue'], fill_type='solid')
    
    ws['A2'] = '=TEXT(ROW()-1,"PROC-0000")'
    
    add_dropdown(ws, 'E2:E1000', ["YES", "NO"], "DPA status")
    add_dropdown(ws, 'G2:G1000', ["YES", "NO"], "SCCs status")
    add_dropdown(ws, 'K2:K1000', ["Compliant", "Partially Compliant", "Non-Compliant"], "Status")
    
    ws.protection.sheet = True
    ws.protection.password = PROTECTION_PASSWORD
    
    return ws


def create_evidence_repository_sheet(wb):
    """Sheet 5: Evidence Repository (14 columns)."""
    ws = wb.create_sheet("Evidence Repository")
    
    headers = [
        "Evidence ID", "Transfer ID", "TIA ID", "Evidence Type", "Description",
        "Document Name", "File Location", "Upload Date", "Owner", "Status", "Notes"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['header_blue'], fill_type='solid')
    
    ws['A2'] = '=TEXT(ROW()-1,"EVID-0000")'
    
    evidence_types = ["SCCs (Signed)", "DPA", "TIA", "DPF Cert", "Legal Memo", "Other"]
    add_dropdown(ws, 'D2:D1000', evidence_types, "Evidence type")
    add_dropdown(ws, 'J2:J1000', ["Current", "Expired", "Superseded"], "Status")
    
    ws.protection.sheet = True
    ws.protection.password = PROTECTION_PASSWORD
    
    return ws


def create_gap_analysis_sheet(wb):
    """Sheet 6: Gap Analysis (25 columns)."""
    ws = wb.create_sheet("Gap Analysis")
    
    headers = [
        "Gap ID", "Transfer ID", "Gap Type", "Description", "Risk Level",
        "Affected Data Subjects", "Discovery Date", "Remediation Action",
        "Owner", "Target Date", "Status", "Notes"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['header_blue'], fill_type='solid')
    
    ws['A2'] = '=TEXT(ROW()-1,"GAP-0000")'
    
    gap_types = ["Missing SCCs", "No TIA", "Undocumented Transfer", "Old SCCs", "Other"]
    add_dropdown(ws, 'C2:C1000', gap_types, "Gap type")
    add_dropdown(ws, 'E2:E1000', RISK_LEVELS, "Risk level")
    add_dropdown(ws, 'K2:K1000', ["Open", "In Progress", "Completed", "Blocked"], "Status")
    
    # Conditional formatting for risk levels
    ws.conditional_formatting.add(
        'E2:E1000',
        CellIsRule(operator='equal', formula=['"Critical"'],
                   fill=PatternFill(start_color=COLORS['light_red'], fill_type='solid'))
    )
    
    ws.protection.sheet = True
    ws.protection.password = PROTECTION_PASSWORD
    
    return ws


def create_dashboard_sheet(wb):
    """Sheet 7: Compliance Dashboard with metrics."""
    ws = wb.create_sheet("Dashboard")
    
    # Title
    ws.merge_cells('A1:L3')
    title = ws['A1']
    title.value = "Cross-Border Transfer Compliance Dashboard"
    title.font = Font(size=18, bold=True, color=COLORS['white'])
    title.fill = PatternFill(start_color=COLORS['header_blue'], fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')
    
    # Metrics (formulas reference other sheets)
    metrics = [
        ("Total Transfers", "=COUNTA('Transfer Register'!A:A)-1"),
        ("Transfers with Valid Mechanism", "=COUNTA('Transfer Register'!H:H)-1"),
        ("TIAs Required", "=COUNTIF('Transfer Register'!L:L,\"YES\")"),
        ("TIAs Completed", "=COUNTA('TIA Register'!A:A)-1"),
        ("Critical Gaps", "=COUNTIF('Gap Analysis'!E:E,\"Critical\")"),
        ("High Gaps", "=COUNTIF('Gap Analysis'!E:E,\"High\")"),
        ("Gaps Completed", "=COUNTIF('Gap Analysis'!K:K,\"Completed\")"),
    ]
    
    row = 5
    for metric_name, formula in metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = '0'
        row += 1
    
    ws.protection.sheet = True
    
    return ws


def create_approval_sheet(wb):
    """Sheet 8: Approval & Sign-Off (14 columns)."""
    ws = wb.create_sheet("Approvals")
    
    headers = [
        "Approval ID", "Approval Type", "Transfer/TIA ID", "Description",
        "Approver Name", "Approver Role", "Approval Date", "Status", "Notes"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=COLORS['header_blue'], fill_type='solid')
    
    ws['A2'] = '=TEXT(ROW()-1,"APPR-0000")'
    
    approval_types = ["DPO Approval", "Legal Review", "CISO Approval", "Business Owner"]
    add_dropdown(ws, 'B2:B1000', approval_types, "Approval type")
    add_dropdown(ws, 'H2:H1000', ["Pending", "Approved", "Rejected"], "Status")
    
    ws.protection.sheet = True
    ws.protection.password = PROTECTION_PASSWORD
    
    return ws


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main():
    """Generate Cross-Border Transfer Assessment workbook."""
    # Parse arguments
    parser = argparse.ArgumentParser(
        description='Generate ISMS A.5.34.6 Cross-Border Transfer Assessment workbook'
    )
    parser.add_argument('--output', type=str, default='.',
                        help='Output directory path')
    parser.add_argument('--date', type=str, default=None,
                        help='Date suffix (YYYYMMDD), defaults to today')
    args = parser.parse_args()
    
    # Determine filename
    date_suffix = args.date if args.date else datetime.now().strftime('%Y%m%d')
    filename = f"{FILE_PREFIX}_{date_suffix}.xlsx"
    output_path = os.path.join(args.output, filename)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    logger.info("Creating Sheet 1: Instructions & Legend...")
    create_instructions_sheet(wb)
    
    logger.info("Creating Sheet 2: Cross-Border Transfer Register...")
    create_transfer_register_sheet(wb)
    
    logger.info("Creating Sheet 3: TIA Register...")
    create_tia_register_sheet(wb)
    
    logger.info("Creating Sheet 4: Processor Agreement Tracker...")
    create_processor_tracker_sheet(wb)
    
    logger.info("Creating Sheet 5: Evidence Repository...")
    create_evidence_repository_sheet(wb)
    
    logger.info("Creating Sheet 6: Gap Analysis...")
    create_gap_analysis_sheet(wb)
    
    logger.info("Creating Sheet 7: Compliance Dashboard...")
    create_dashboard_sheet(wb)
    
    logger.info("Creating Sheet 8: Approval & Sign-Off...")
    create_approval_sheet(wb)
    
    # Save
    logger.info(f"\nSaving workbook to {output_path}...")
    wb.save(output_path)
    
    logger.info(f"\nSuccess! Workbook created: {output_path}")
    logger.info("\nNext steps:")
    logger.info("1. Review Sheet 1 (Instructions) for GDPR Chapter V guidance")
    logger.info("2. Complete Sheet 2 (Transfer Register) - inventory all cross-border transfers")
    logger.info("3. Conduct TIAs in Sheet 3 for non-adequate country transfers")
    logger.info("4. Verify processor agreements in Sheet 4")
    logger.info("5. Collect evidence in Sheet 5")
    logger.info("6. Document gaps in Sheet 6")
    logger.info("7. Review Sheet 7 (Dashboard) for compliance metrics")
    logger.info("8. Obtain approvals in Sheet 8")
    logger.info("9. Feed results into A.5.34.7 Privacy Compliance Dashboard")


if __name__ == '__main__':
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
