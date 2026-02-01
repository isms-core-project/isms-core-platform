#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.12-13.4 - Classification and Labelling Compliance Dashboard
================================================================================

ISO/IEC 27001:2022 Controls A.5.12 & A.5.13
Assessment Domain 4 of 4: Compliance Dashboard

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive compliance dashboard for monitoring
the organization's classification and labelling program effectiveness.

**Purpose:**
Provide executive-level visibility into classification scheme adoption,
labelling compliance, and program maturity.

**Generated Workbook Structure:**
1. Executive_Summary - Key metrics and compliance status
2. Compliance_Metrics - Detailed KPIs and measurements
3. Control_Assessment - A.5.12 and A.5.13 control evaluation
4. Maturity_Assessment - Program maturity scoring
5. Risk_Register - Classification-related risks
6. Remediation_Tracker - Action items and progress
7. Evidence_Register - Audit evidence tracking
8. Approval_SignOff - Dashboard review and approval

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.12-13.4"
WORKBOOK_NAME = "Classification and Labelling Compliance Dashboard"
CONTROL_ID = "A.5.12-13"
CONTROL_NAME = "Classification and Labelling of Information"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "metric_good": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "metric_warning": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "metric_bad": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "border": border_thin,
    }


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    for attr, value in style_dict.items():
        if attr != "border":
            setattr(cell, attr, value)


def set_column_widths(ws, widths):
    """Set column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================
def create_executive_summary_sheet(ws, styles):
    """Create the Executive Summary sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{DOCUMENT_ID} - Executive Summary"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    apply_style(ws["A2"], styles["subheader"])

    # Overall Status
    ws["A4"] = "OVERALL COMPLIANCE STATUS"
    ws["A4"].font = Font(bold=True, size=12)

    status_items = [
        ("Classification Scheme Defined", ""),
        ("Labelling Procedures Documented", ""),
        ("Asset Inventory Coverage", ""),
        ("Staff Awareness Training", ""),
        ("Automated Labelling Deployed", ""),
        ("Periodic Review Conducted", ""),
    ]

    for i, (item, status) in enumerate(status_items, start=5):
        ws.cell(row=i, column=1, value=item).border = styles["border"]
        cell = ws.cell(row=i, column=2, value=status)
        cell.border = styles["border"]
        cell.fill = styles["input_cell"]["fill"]

    # Key Metrics
    ws["A13"] = "KEY METRICS"
    ws["A13"].font = Font(bold=True, size=12)

    metrics_headers = ["Metric", "Value", "Target", "Status"]
    for col, header in enumerate(metrics_headers, start=1):
        cell = ws.cell(row=14, column=col, value=header)
        apply_style(cell, styles["column_header"])

    metrics = [
        ("Total Assets Classified", "", "100%", ""),
        ("Assets Properly Labelled", "", "95%", ""),
        ("Classification Review Currency", "", "100%", ""),
        ("Labelling Tool Coverage", "", "90%", ""),
        ("Training Completion Rate", "", "100%", ""),
        ("Exception Requests (Open)", "", "<5", ""),
    ]

    for i, (metric, value, target, status) in enumerate(metrics, start=15):
        ws.cell(row=i, column=1, value=metric).border = styles["border"]
        for col in [2, 3, 4]:
            cell = ws.cell(row=i, column=col, value=[value, target, status][col-2])
            cell.border = styles["border"]
            if col in [2, 4]:
                cell.fill = styles["input_cell"]["fill"]

    # Assessment Period
    ws["A23"] = "ASSESSMENT PERIOD"
    ws["A23"].font = Font(bold=True, size=12)

    ws["A24"] = "Review Date:"
    ws["B24"] = GENERATED_DATE
    ws["A25"] = "Review Period:"
    ws["B25"] = ""
    ws["A26"] = "Next Review:"
    ws["B26"] = ""
    ws["A27"] = "Reviewer:"
    ws["B27"] = ""

    for row in range(24, 28):
        ws[f"A{row}"].font = Font(bold=True)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]

    # Data validation for status
    dv_status = DataValidation(
        type="list",
        formula1='"✅ Complete,⚠️ In Progress,❌ Not Started,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("B5:B10")
    dv_status.add("D15:D20")

    set_column_widths(ws, {"A": 35, "B": 20, "C": 15, "D": 18})
    logger.info("Created Executive_Summary sheet")


def create_compliance_metrics_sheet(ws, styles):
    """Create the Compliance Metrics sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Classification and Labelling Compliance Metrics"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = "Detailed KPIs for monitoring program effectiveness"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Metric ID", "Metric Name", "Description", "Measurement Method",
        "Target", "Current", "Trend", "Owner", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    metrics = [
        ["M-CL-001", "Classification Coverage", "% of assets with assigned classification",
         "Asset inventory count", "100%", "", "", "InfoSec", ""],
        ["M-CL-002", "Labelling Compliance", "% of classified assets properly labelled",
         "Labelling audit", "95%", "", "", "InfoSec", ""],
        ["M-CL-003", "Classification Currency", "% of assets reviewed within policy period",
         "Review date tracking", "100%", "", "", "Data Gov", ""],
        ["M-CL-004", "Reclassification Timeliness", "Avg days to process reclassification requests",
         "Request tracking", "<5 days", "", "", "InfoSec", ""],
        ["M-CL-005", "Training Completion", "% of staff completing classification training",
         "LMS records", "100%", "", "", "HR/InfoSec", ""],
        ["M-CL-006", "Automation Coverage", "% of asset types with automated labelling",
         "Tool inventory", "90%", "", "", "IT Ops", ""],
        ["M-CL-007", "Exception Rate", "% of assets with classification exceptions",
         "Exception register", "<2%", "", "", "InfoSec", ""],
        ["M-CL-008", "Incident Rate", "Classification-related security incidents",
         "Incident tracking", "0/quarter", "", "", "InfoSec", ""],
        ["M-CL-009", "Audit Findings", "Open audit findings related to classification",
         "Audit reports", "0", "", "", "Compliance", ""],
        ["M-CL-010", "Policy Acknowledgment", "% of staff acknowledging classification policy",
         "HR records", "100%", "", "", "HR", ""],
    ]

    for row_idx, metric in enumerate(metrics, start=5):
        for col_idx, value in enumerate(metric, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx in [6, 7, 9]:  # Current, Trend, Status columns
                cell.fill = styles["input_cell"]["fill"]

    dv_trend = DataValidation(
        type="list",
        formula1='"↑ Improving,→ Stable,↓ Declining,New"',
        allow_blank=True
    )
    ws.add_data_validation(dv_trend)
    dv_trend.add("G5:G30")

    dv_status = DataValidation(
        type="list",
        formula1='"✅ On Target,⚠️ At Risk,❌ Below Target,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("I5:I30")

    set_column_widths(ws, {
        "A": 12, "B": 22, "C": 35, "D": 22, "E": 12,
        "F": 12, "G": 12, "H": 12, "I": 15
    })
    logger.info("Created Compliance_Metrics sheet")


def create_control_assessment_sheet(ws, styles):
    """Create the Control Assessment sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Control A.5.12 & A.5.13 Assessment"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Detailed evaluation of ISO 27001:2022 control requirements"
    apply_style(ws["A2"], styles["subheader"])

    # A.5.12 Assessment
    ws["A4"] = "CONTROL A.5.12: CLASSIFICATION OF INFORMATION"
    ws["A4"].font = Font(bold=True, size=12)

    a512_headers = ["Requirement", "Implementation", "Evidence", "Gap", "Score", "Status"]
    for col, header in enumerate(a512_headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["column_header"])

    a512_reqs = [
        ["Classification scheme defined based on CIA needs", "", "", "", "", ""],
        ["Classification considers legal/regulatory requirements", "", "", "", "", ""],
        ["Information owners assigned and trained", "", "", "", "", ""],
        ["Classification review process established", "", "", "", "", ""],
        ["Reclassification procedures documented", "", "", "", "", ""],
        ["Cross-organization classification alignment", "", "", "", "", ""],
    ]

    for row_idx, req in enumerate(a512_reqs, start=6):
        for col_idx, value in enumerate(req, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx > 1:
                cell.fill = styles["input_cell"]["fill"]

    # A.5.13 Assessment
    ws["A14"] = "CONTROL A.5.13: LABELLING OF INFORMATION"
    ws["A14"].font = Font(bold=True, size=12)

    for col, header in enumerate(a512_headers, start=1):
        cell = ws.cell(row=15, column=col, value=header)
        apply_style(cell, styles["column_header"])

    a513_reqs = [
        ["Labelling procedures reflect classification scheme", "", "", "", "", ""],
        ["Labels applied to physical and digital assets", "", "", "", "", ""],
        ["Labels easily recognizable", "", "", "", "", ""],
        ["Metadata tagging implemented", "", "", "", "", ""],
        ["Labelling guidance for transmission", "", "", "", "", ""],
        ["Staff trained on labelling procedures", "", "", "", "", ""],
    ]

    for row_idx, req in enumerate(a513_reqs, start=16):
        for col_idx, value in enumerate(req, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx > 1:
                cell.fill = styles["input_cell"]["fill"]

    dv_score = DataValidation(
        type="list",
        formula1='"5 - Optimized,4 - Managed,3 - Defined,2 - Developing,1 - Initial,0 - Non-existent"',
        allow_blank=True
    )
    ws.add_data_validation(dv_score)
    dv_score.add("E6:E11")
    dv_score.add("E16:E21")

    dv_status = DataValidation(
        type="list",
        formula1='"✅ Compliant,⚠️ Partial,❌ Non-Compliant,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("F6:F11")
    dv_status.add("F16:F21")

    set_column_widths(ws, {
        "A": 45, "B": 30, "C": 25, "D": 25, "E": 18, "F": 15
    })
    logger.info("Created Control_Assessment sheet")


def create_maturity_assessment_sheet(ws, styles):
    """Create the Maturity Assessment sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "Classification Program Maturity Assessment"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "CMMI-based maturity model for classification and labelling"
    apply_style(ws["A2"], styles["subheader"])

    # Maturity scale reference
    ws["A4"] = "MATURITY SCALE REFERENCE"
    ws["A4"].font = Font(bold=True, size=11)

    scale = [
        ("5 - Optimized", "Continuous improvement, automated, metrics-driven"),
        ("4 - Managed", "Measured and controlled, consistent performance"),
        ("3 - Defined", "Documented, standardized across organization"),
        ("2 - Developing", "Reactive, informal, department-specific"),
        ("1 - Initial", "Ad hoc, unpredictable, poorly controlled"),
        ("0 - Non-existent", "No awareness, no processes"),
    ]

    for i, (level, desc) in enumerate(scale, start=5):
        ws.cell(row=i, column=1, value=level).font = Font(bold=True)
        ws.cell(row=i, column=1).border = styles["border"]
        cell = ws.cell(row=i, column=2, value=desc)
        cell.border = styles["border"]
        ws.merge_cells(f"B{i}:D{i}")

    # Domain assessment
    ws["A13"] = "DOMAIN MATURITY ASSESSMENT"
    ws["A13"].font = Font(bold=True, size=11)

    domain_headers = ["Domain", "Current Level", "Target Level", "Gap", "Priority", "Notes"]
    for col, header in enumerate(domain_headers, start=1):
        cell = ws.cell(row=14, column=col, value=header)
        apply_style(cell, styles["column_header"])

    domains = [
        ["Policy & Governance", "", "", "", "", ""],
        ["Classification Scheme", "", "", "", "", ""],
        ["Labelling Standards", "", "", "", "", ""],
        ["Asset Inventory", "", "", "", "", ""],
        ["Automation & Tools", "", "", "", "", ""],
        ["Training & Awareness", "", "", "", "", ""],
        ["Monitoring & Metrics", "", "", "", "", ""],
        ["Incident Response", "", "", "", "", ""],
        ["OVERALL MATURITY", "", "", "", "", ""],
    ]

    for row_idx, domain in enumerate(domains, start=15):
        for col_idx, value in enumerate(domain, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            if col_idx > 1:
                cell.fill = styles["input_cell"]["fill"]
        if domain[0] == "OVERALL MATURITY":
            ws.cell(row=row_idx, column=1).font = Font(bold=True)

    dv_level = DataValidation(
        type="list",
        formula1='"5 - Optimized,4 - Managed,3 - Defined,2 - Developing,1 - Initial,0 - Non-existent"',
        allow_blank=True
    )
    ws.add_data_validation(dv_level)
    dv_level.add("B15:C23")

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_priority)
    dv_priority.add("E15:E23")

    set_column_widths(ws, {
        "A": 25, "B": 18, "C": 18, "D": 10, "E": 12, "F": 35
    })
    logger.info("Created Maturity_Assessment sheet")


def create_risk_register_sheet(ws, styles):
    """Create the Risk Register sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Classification-Related Risk Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Track risks related to information classification and labelling"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Risk ID", "Risk Description", "Risk Category", "Likelihood",
        "Impact", "Risk Score", "Mitigation", "Owner", "Status", "Review Date"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    sample_risks = [
        ["RSK-CL-001", "Unclassified sensitive data exposed", "Data Breach",
         "", "", "", "Implement auto-classification", "InfoSec", "", ""],
        ["RSK-CL-002", "Inconsistent labelling leads to mishandling", "Compliance",
         "", "", "", "Standardize labelling tools", "IT Ops", "", ""],
        ["RSK-CL-003", "Staff unaware of classification requirements", "Human Error",
         "", "", "", "Mandatory training program", "HR", "", ""],
        ["RSK-CL-004", "Classification not reviewed as data ages", "Data Lifecycle",
         "", "", "", "Automated review reminders", "Data Gov", "", ""],
    ]

    for row_idx, risk in enumerate(sample_risks, start=5):
        for col_idx, value in enumerate(risk, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx in [4, 5, 6, 9, 10]:
                cell.fill = styles["input_cell"]["fill"]

    for row in range(9, 25):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_likelihood = DataValidation(
        type="list",
        formula1='"5 - Almost Certain,4 - Likely,3 - Possible,2 - Unlikely,1 - Rare"',
        allow_blank=True
    )
    ws.add_data_validation(dv_likelihood)
    dv_likelihood.add("D5:D30")

    dv_impact = DataValidation(
        type="list",
        formula1='"5 - Catastrophic,4 - Major,3 - Moderate,2 - Minor,1 - Insignificant"',
        allow_blank=True
    )
    ws.add_data_validation(dv_impact)
    dv_impact.add("E5:E30")

    dv_status = DataValidation(
        type="list",
        formula1='"✅ Mitigated,⚠️ In Progress,❌ Open,Accepted,Transferred"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("I5:I30")

    set_column_widths(ws, {
        "A": 12, "B": 35, "C": 15, "D": 18, "E": 18,
        "F": 12, "G": 30, "H": 12, "I": 15, "J": 12
    })
    logger.info("Created Risk_Register sheet")


def create_remediation_tracker_sheet(ws, styles):
    """Create the Remediation Tracker sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Remediation Action Tracker"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = "Track remediation actions for classification and labelling gaps"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Action ID", "Description", "Source", "Priority",
        "Owner", "Due Date", "Progress", "Status", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    for row in range(5, 30):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_source = DataValidation(
        type="list",
        formula1='"Audit Finding,Risk Assessment,Gap Analysis,Incident,Self-Assessment,Regulatory,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_source)
    dv_source.add("C5:C40")

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_priority)
    dv_priority.add("D5:D40")

    dv_progress = DataValidation(
        type="list",
        formula1='"0%,25%,50%,75%,100%"',
        allow_blank=True
    )
    ws.add_data_validation(dv_progress)
    dv_progress.add("G5:G40")

    dv_status = DataValidation(
        type="list",
        formula1='"✅ Complete,⚠️ In Progress,❌ Not Started,On Hold,Cancelled"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H40")

    set_column_widths(ws, {
        "A": 12, "B": 40, "C": 18, "D": 12, "E": 15,
        "F": 12, "G": 12, "H": 15, "I": 30
    })
    logger.info("Created Remediation_Tracker sheet")


def create_evidence_register_sheet(ws, styles):
    """Create the Evidence Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Evidence Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Track audit evidence for classification and labelling compliance"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Evidence ID", "Description", "Evidence Type",
        "Related Control", "Location", "Collected Date",
        "Collected By", "Verification Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    for row in range(5, 25):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_type = DataValidation(
        type="list",
        formula1='"Policy document,Procedure,Configuration,Screenshot,Training record,Audit report,Assessment,Metrics report,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C5:C30")

    dv_control = DataValidation(
        type="list",
        formula1='"A.5.12,A.5.13,Both"',
        allow_blank=True
    )
    ws.add_data_validation(dv_control)
    dv_control.add("D5:D30")

    dv_status = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending Review,❌ Not Verified,Expired"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H30")

    set_column_widths(ws, {
        "A": 15, "B": 40, "C": 20, "D": 15, "E": 30, "F": 15, "G": 15, "H": 18
    })
    logger.info("Created Evidence_Register sheet")


def create_approval_sheet(ws, styles):
    """Create the Approval Sign-Off sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "Dashboard Review and Approval"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "Formal approval of compliance dashboard review"
    apply_style(ws["A2"], styles["subheader"])

    ws["A4"] = "Review Period:"
    ws["B4"] = ""
    ws["A5"] = "Overall Compliance:"
    ws["B5"] = ""
    ws["A6"] = "Maturity Level:"
    ws["B6"] = ""
    ws["A7"] = "Open Risks:"
    ws["B7"] = ""

    for row in range(4, 8):
        ws[f"A{row}"].font = Font(bold=True)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]

    ws["A10"] = "APPROVAL SIGNATURES"
    ws["A10"].font = Font(bold=True, size=12)

    approval_headers = ["Role", "Name", "Signature", "Date", "Decision", "Comments"]
    for col, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=12, column=col, value=header)
        apply_style(cell, styles["column_header"])

    approvers = [
        "Chief Information Security Officer",
        "Data Protection Officer",
        "IT Director",
        "Compliance Officer",
        "Internal Audit",
    ]

    for row_idx, role in enumerate(approvers, start=13):
        ws.cell(row=row_idx, column=1, value=role).border = styles["border"]
        for col in range(2, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected,Pending"',
        allow_blank=True
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add("E13:E20")

    set_column_widths(ws, {"A": 35, "B": 25, "C": 20, "D": 15, "E": 22, "F": 30})
    logger.info("Created Approval_SignOff sheet")


# =============================================================================
# MAIN FUNCTION
# =============================================================================
def main():
    """Main entry point."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS-IMP-A.5.12-13.4 Compliance Dashboard Generator")
        logger.info("=" * 70)

        wb = Workbook()
        styles = setup_styles()

        ws_exec = wb.active
        ws_exec.title = "Executive_Summary"

        ws_metrics = wb.create_sheet("Compliance_Metrics")
        ws_control = wb.create_sheet("Control_Assessment")
        ws_maturity = wb.create_sheet("Maturity_Assessment")
        ws_risk = wb.create_sheet("Risk_Register")
        ws_remediation = wb.create_sheet("Remediation_Tracker")
        ws_evidence = wb.create_sheet("Evidence_Register")
        ws_approval = wb.create_sheet("Approval_SignOff")

        create_executive_summary_sheet(ws_exec, styles)
        create_compliance_metrics_sheet(ws_metrics, styles)
        create_control_assessment_sheet(ws_control, styles)
        create_maturity_assessment_sheet(ws_maturity, styles)
        create_risk_register_sheet(ws_risk, styles)
        create_remediation_tracker_sheet(ws_remediation, styles)
        create_evidence_register_sheet(ws_evidence, styles)
        create_approval_sheet(ws_approval, styles)

        wb.save(OUTPUT_FILENAME)
        logger.info("=" * 70)
        logger.info("SUCCESS: Workbook saved as %s", OUTPUT_FILENAME)
        logger.info("=" * 70)
        return 0

    except Exception as e:
        logger.error("FAILED: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.12-13 Classification & Labelling
# =============================================================================
