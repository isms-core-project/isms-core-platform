#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.12-13.3 - Asset Classification Inventory Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.5.12 & A.5.13
Assessment Domain 3 of 4: Asset Classification Inventory

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel workbook for tracking the
classification status of information assets across the organization.

**Purpose:**
Maintain a complete inventory of information assets with their assigned
classification levels, owners, and compliance status.

**Generated Workbook Structure:**
1. Instructions - Guidance on asset classification tracking
2. Asset_Inventory - Master list of classified assets
3. Classification_Summary - Statistics by level and category
4. Reclassification_Log - Track classification changes
5. Gap_Analysis - Unclassified or improperly classified assets
6. Evidence_Register - Audit evidence tracking
7. Approval_SignOff - Inventory approval

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.12-13.3"
WORKBOOK_NAME = "Asset Classification Inventory"
CONTROL_ID = "A.5.12-13"
CONTROL_NAME = "Classification and Labelling of Information"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "level_restricted": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        "level_confidential": PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid"),
        "level_internal": PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid"),
        "level_public": PatternFill(start_color="74C0FC", end_color="74C0FC", fill_type="solid"),
    }


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    for attr, value in style_dict.items():
        if attr != "border":
            setattr(cell, attr, value)


def set_column_widths(ws, widths):
    """Set column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    apply_style(ws["A2"], styles["subheader"])

    instructions = [
        "",
        "PURPOSE",
        "This workbook maintains an inventory of all classified information assets",
        "to ensure complete coverage of the classification scheme and enable",
        "compliance tracking and reporting.",
        "",
        "INVENTORY SCOPE",
        "• Digital documents (files, databases, applications)",
        "• Physical documents and records",
        "• Data repositories and storage systems",
        "• Information systems and applications",
        "• Communication channels and email",
        "",
        "CLASSIFICATION OWNERSHIP",
        "• The information owner is responsible for assigning classification",
        "• Custodians must apply appropriate labels and controls",
        "• Users must handle information according to its classification",
        "",
        "HOW TO USE THIS WORKBOOK",
        "1. Add assets to Asset_Inventory with classification details",
        "2. Review Classification_Summary for statistics and gaps",
        "3. Log any reclassifications in Reclassification_Log",
        "4. Address items in Gap_Analysis requiring attention",
        "5. Track evidence in Evidence_Register",
        "6. Obtain approval in Approval_SignOff",
        "",
        "REVIEW CYCLE",
        "• Full inventory review: Annual",
        "• Classification verification: Upon change or quarterly",
        "• New assets: Classify before deployment",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, line in enumerate(instructions, start=4):
        ws[f"A{i}"] = line
        if line in ["PURPOSE", "INVENTORY SCOPE", "CLASSIFICATION OWNERSHIP",
                    "HOW TO USE THIS WORKBOOK", "REVIEW CYCLE"]:
            ws[f"A{i}"].font = Font(bold=True, size=11)

    set_column_widths(ws, {"A": 100})
    logger.info("Created Instructions sheet")


def create_asset_inventory_sheet(ws, styles):
    """Create the Asset Inventory sheet."""
    ws.merge_cells("A1:N1")
    ws["A1"] = "Information Asset Classification Inventory"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:N2")
    ws["A2"] = "Master list of classified information assets"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Asset ID", "Asset Name", "Asset Type", "Description",
        "Classification", "Owner", "Custodian", "Location/System",
        "Labelling Status", "Last Review", "Next Review",
        "Regulatory Req", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add sample data rows
    sample_assets = [
        ["AST-001", "Customer Database", "Database", "Production customer records",
         "RESTRICTED", "CRM Manager", "DBA Team", "SQL Server Prod",
         "✅ Labelled", "2026-01-15", "2026-07-15", "GDPR, nDSG", ""],
        ["AST-002", "Financial Reports", "Document Set", "Quarterly financial statements",
         "CONFIDENTIAL", "CFO", "Finance Team", "SharePoint Finance",
         "✅ Labelled", "2026-01-10", "2026-04-10", "SOX", ""],
        ["AST-003", "Employee Handbook", "Document", "HR policies and procedures",
         "INTERNAL", "HR Director", "HR Team", "Intranet",
         "✅ Labelled", "2025-12-01", "2026-12-01", "", ""],
        ["AST-004", "Marketing Materials", "Document Set", "Public brochures and website",
         "PUBLIC", "Marketing Lead", "Marketing", "Website/Drive",
         "✅ Labelled", "2025-11-15", "2026-11-15", "", ""],
    ]

    for row_idx, asset in enumerate(sample_assets, start=5):
        for col_idx, value in enumerate(asset, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")

            # Color-code classification column
            if col_idx == 5:
                if value == "RESTRICTED":
                    cell.fill = styles["level_restricted"]
                    cell.font = Font(bold=True, color="FFFFFF")
                elif value == "CONFIDENTIAL":
                    cell.fill = styles["level_confidential"]
                    cell.font = Font(bold=True)
                elif value == "INTERNAL":
                    cell.fill = styles["level_internal"]
                    cell.font = Font(bold=True)
                elif value == "PUBLIC":
                    cell.fill = styles["level_public"]
                    cell.font = Font(bold=True)

    # Add empty input rows
    for row in range(9, 50):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    # Data validations
    dv_type = DataValidation(
        type="list",
        formula1='"Database,Document,Document Set,Application,System,Repository,Email,Media,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C5:C100")

    dv_class = DataValidation(
        type="list",
        formula1='"RESTRICTED,CONFIDENTIAL,INTERNAL,PUBLIC"',
        allow_blank=True
    )
    ws.add_data_validation(dv_class)
    dv_class.add("E5:E100")

    dv_label = DataValidation(
        type="list",
        formula1='"✅ Labelled,⚠️ Partial,❌ Not Labelled,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_label)
    dv_label.add("I5:I100")

    set_column_widths(ws, {
        "A": 12, "B": 25, "C": 15, "D": 30, "E": 15,
        "F": 18, "G": 15, "H": 20, "I": 15,
        "J": 12, "K": 12, "L": 15, "M": 25
    })
    logger.info("Created Asset_Inventory sheet")


def create_classification_summary_sheet(ws, styles):
    """Create the Classification Summary sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Classification Summary Dashboard"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Statistics and metrics by classification level and asset type"
    apply_style(ws["A2"], styles["subheader"])

    # By Classification Level
    ws["A4"] = "ASSETS BY CLASSIFICATION LEVEL"
    ws["A4"].font = Font(bold=True, size=12)

    level_headers = ["Classification", "Count", "Percentage", "Labelled", "Unlabelled", "Compliance"]
    for col, header in enumerate(level_headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["column_header"])

    levels = [
        ["RESTRICTED", "", "", "", "", ""],
        ["CONFIDENTIAL", "", "", "", "", ""],
        ["INTERNAL", "", "", "", "", ""],
        ["PUBLIC", "", "", "", "", ""],
        ["UNCLASSIFIED", "", "", "", "", ""],
        ["TOTAL", "", "100%", "", "", ""],
    ]

    for row_idx, level in enumerate(levels, start=6):
        for col_idx, value in enumerate(level, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            if col_idx > 1:
                cell.fill = styles["input_cell"]["fill"]
                cell.alignment = Alignment(horizontal="center")

    # By Asset Type
    ws["A14"] = "ASSETS BY TYPE"
    ws["A14"].font = Font(bold=True, size=12)

    type_headers = ["Asset Type", "Total", "RESTRICTED", "CONFIDENTIAL", "INTERNAL", "PUBLIC"]
    for col, header in enumerate(type_headers, start=1):
        cell = ws.cell(row=15, column=col, value=header)
        apply_style(cell, styles["column_header"])

    asset_types = [
        "Database", "Document", "Document Set", "Application",
        "System", "Repository", "Email", "Media", "Other", "TOTAL"
    ]

    for row_idx, asset_type in enumerate(asset_types, start=16):
        ws.cell(row=row_idx, column=1, value=asset_type).border = styles["border"]
        for col in range(2, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]
            cell.alignment = Alignment(horizontal="center")

    # By Department
    ws["A28"] = "ASSETS BY DEPARTMENT/OWNER"
    ws["A28"].font = Font(bold=True, size=12)

    dept_headers = ["Department", "Total", "RESTRICTED", "CONFIDENTIAL", "INTERNAL", "PUBLIC", "Compliance %"]
    for col, header in enumerate(dept_headers, start=1):
        cell = ws.cell(row=29, column=col, value=header)
        apply_style(cell, styles["column_header"])

    for row in range(30, 40):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    set_column_widths(ws, {
        "A": 20, "B": 12, "C": 15, "D": 15, "E": 15, "F": 15, "G": 15
    })
    logger.info("Created Classification_Summary sheet")


def create_reclassification_log_sheet(ws, styles):
    """Create the Reclassification Log sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Reclassification Change Log"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Track all classification level changes with justification and approval"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Change ID", "Asset ID", "Asset Name", "Previous Class",
        "New Class", "Reason for Change", "Requested By",
        "Approved By", "Change Date", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    for row in range(5, 35):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_class = DataValidation(
        type="list",
        formula1='"RESTRICTED,CONFIDENTIAL,INTERNAL,PUBLIC"',
        allow_blank=True
    )
    ws.add_data_validation(dv_class)
    dv_class.add("D5:D50")
    dv_class.add("E5:E50")

    dv_reason = DataValidation(
        type="list",
        formula1='"Value change,Regulatory requirement,Business need,Data lifecycle,Merger/divestiture,Error correction,Periodic review,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_reason)
    dv_reason.add("F5:F50")

    dv_status = DataValidation(
        type="list",
        formula1='"✅ Complete,⚠️ Pending Approval,❌ Rejected,In Progress"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("J5:J50")

    set_column_widths(ws, {
        "A": 12, "B": 12, "C": 25, "D": 15, "E": 15,
        "F": 25, "G": 18, "H": 18, "I": 12, "J": 18
    })
    logger.info("Created Reclassification_Log sheet")


def create_gap_analysis_sheet(ws, styles):
    """Create the Gap Analysis sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Classification Gap Analysis"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = "Identify assets requiring classification attention"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Gap ID", "Asset/Area", "Gap Type", "Description",
        "Risk Level", "Remediation Action", "Owner",
        "Due Date", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    sample_gaps = [
        ["GAP-001", "Legacy File Server", "Unclassified Assets",
         "File server contains unclassified documents from pre-policy era",
         "High", "Conduct classification sweep of all files",
         "IT Manager", "2026-03-31", ""],
        ["GAP-002", "Email Archives", "Incomplete Labelling",
         "Historical emails lack classification metadata",
         "Medium", "Apply auto-classification rules to archive",
         "Email Admin", "2026-04-30", ""],
        ["GAP-003", "Third-party SaaS", "No Labelling Capability",
         "SaaS application doesn't support sensitivity labels",
         "Medium", "Implement compensating controls (DLP)",
         "SaaS Owner", "2026-05-31", ""],
    ]

    for row_idx, gap in enumerate(sample_gaps, start=5):
        for col_idx, value in enumerate(gap, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx == 9:
                cell.fill = styles["input_cell"]["fill"]

    for row in range(8, 30):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_type = DataValidation(
        type="list",
        formula1='"Unclassified Assets,Incomplete Labelling,Misclassification,No Labelling Capability,Inconsistent Labels,Missing Metadata,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C5:C50")

    dv_risk = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_risk)
    dv_risk.add("E5:E50")

    dv_status = DataValidation(
        type="list",
        formula1='"✅ Resolved,⚠️ In Progress,❌ Open,Accepted"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("I5:I50")

    set_column_widths(ws, {
        "A": 12, "B": 20, "C": 22, "D": 45, "E": 12,
        "F": 40, "G": 15, "H": 12, "I": 15
    })
    logger.info("Created Gap_Analysis sheet")


def create_evidence_register_sheet(ws, styles):
    """Create the Evidence Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Evidence Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Track audit evidence supporting asset classification"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Evidence ID", "Description", "Evidence Type",
        "Related Asset/Gap", "Location", "Collected Date",
        "Collected By", "Verification Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    for row in range(5, 25):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_type = DataValidation(
        type="list",
        formula1='"Inventory export,Classification report,Label screenshot,Audit finding,Training record,Policy acknowledgment,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C5:C30")

    dv_status = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending Review,❌ Not Verified,Expired"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H30")

    set_column_widths(ws, {
        "A": 15, "B": 40, "C": 22, "D": 20, "E": 30, "F": 15, "G": 15, "H": 18
    })
    logger.info("Created Evidence_Register sheet")


def create_approval_sheet(ws, styles):
    """Create the Approval Sign-Off sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "Asset Classification Inventory Approval"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "Formal approval record for classification inventory"
    apply_style(ws["A2"], styles["subheader"])

    ws["A4"] = "Inventory Version:"
    ws["B4"] = "1.0"
    ws["A5"] = "Review Period:"
    ws["B5"] = ""
    ws["A6"] = "Total Assets:"
    ws["B6"] = ""
    ws["A7"] = "Compliance Rate:"
    ws["B7"] = ""

    for row in range(4, 8):
        ws[f"A{row}"].font = Font(bold=True)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]

    ws["A10"] = "APPROVAL SIGNATURES"
    ws["A10"].font = Font(bold=True, size=12)

    approval_headers = ["Role", "Name", "Signature", "Date", "Decision", "Comments"]
    for col, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=12, column=col, value=header)
        apply_style(cell, styles["column_header"])

    approvers = [
        "Information Security Officer",
        "Data Governance Lead",
        "IT Asset Manager",
        "Compliance Officer",
    ]

    for row_idx, role in enumerate(approvers, start=13):
        ws.cell(row=row_idx, column=1, value=role).border = styles["border"]
        for col in range(2, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected,Pending"',
        allow_blank=True
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add("E13:E20")

    set_column_widths(ws, {"A": 30, "B": 25, "C": 20, "D": 15, "E": 22, "F": 30})
    logger.info("Created Approval_SignOff sheet")


# =============================================================================
# MAIN FUNCTION
# =============================================================================
def main():
    """Main entry point."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS-IMP-A.5.12-13.3 Asset Classification Inventory Generator")
        logger.info("=" * 70)

        wb = Workbook()
        styles = setup_styles()

        ws_instructions = wb.active
        ws_instructions.title = "Instructions"

        ws_inventory = wb.create_sheet("Asset_Inventory")
        ws_summary = wb.create_sheet("Classification_Summary")
        ws_reclass = wb.create_sheet("Reclassification_Log")
        ws_gaps = wb.create_sheet("Gap_Analysis")
        ws_evidence = wb.create_sheet("Evidence_Register")
        ws_approval = wb.create_sheet("Approval_SignOff")

        create_instructions_sheet(ws_instructions, styles)
        create_asset_inventory_sheet(ws_inventory, styles)
        create_classification_summary_sheet(ws_summary, styles)
        create_reclassification_log_sheet(ws_reclass, styles)
        create_gap_analysis_sheet(ws_gaps, styles)
        create_evidence_register_sheet(ws_evidence, styles)
        create_approval_sheet(ws_approval, styles)

        wb.save(OUTPUT_FILENAME)
        logger.info("=" * 70)
        logger.info("SUCCESS: Workbook saved as %s", OUTPUT_FILENAME)
        logger.info("=" * 70)
        return 0

    except Exception as e:
        logger.error("FAILED: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.12-13 Classification & Labelling
# =============================================================================
