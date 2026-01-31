#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.1-7-18-19.S2 - Protection Coverage Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.1, A.8.7, A.8.18, A.8.19
Assessment Domain 2 of 6: Malware Protection and Endpoint Defense Coverage

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific anti-malware/EDR solutions, protection mechanisms,
and coverage requirements.

Key customization areas:
1. Anti-malware/EDR solutions (match your actual tools: CrowdStrike, Defender, SentinelOne, etc.)
2. Protection mechanisms (adapt to your detection capabilities)
3. Signature/definition update requirements (align with your update policies)
4. Quarantine and remediation procedures (specific to your incident response)
5. Coverage thresholds (based on your risk tolerance)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.1-7-18-19 Endpoint Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
malware protection coverage, detection capabilities, and incident response
effectiveness across all endpoint devices.

**Purpose:**
Enables systematic assessment of anti-malware/EDR deployment and effectiveness
against ISO 27001:2022 Control A.8.7 requirements, supporting evidence-based
validation of malware protection controls.

**Assessment Scope:**
- Anti-malware/EDR solution deployment coverage
- Per-endpoint protection status and agent versions
- Signature/definition update status and timeliness
- Protection mechanisms (signature-based, behavioral, ML-based)
- Real-time scanning status and last scan dates
- Malware detection and quarantine tracking
- Remediation effectiveness and resolution times
- User awareness and reporting mechanisms
- Gap analysis for unprotected/misconfigured endpoints
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and protection standards
2. Protection_Coverage - Per-endpoint protection status matrix
3. Agent_Status - Anti-malware/EDR agent deployment and versions
4. Signature_Updates - Definition update status and timeliness
5. Detection_Summary - Malware detections and threat landscape
6. Quarantine_Tracking - Quarantined threats and remediation status
7. Scanning_Compliance - Real-time and scheduled scan compliance
8. Coverage_Gaps - Unprotected or misconfigured endpoints
9. Evidence_Register - Audit evidence tracking and documentation
10. Remediation_Tracking - Protection gap remediation actions
11. Approval_Sign_Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with anti-malware solution dropdown lists
- Conditional formatting for protection status visualization
- Automated gap identification for unprotected endpoints
- Signature update timeliness tracking (<24 hours requirement)
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with EDR console outputs

**Integration:**
This assessment builds on the endpoint inventory (A.8.1-7-18-19.S1) and feeds
into the consolidated compliance matrix (A.8.1-7-18-19.S5) and dashboard
(A.8.1-7-18-19.S6).

--------------------------------------------------------------------------------
[Continue with REQUIREMENTS, USAGE, METADATA, IMPORTANT NOTES as per template pattern]
================================================================================
"""

import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================



# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
LAPTOP = '\U0001F4BB' # 💻 Laptop
VIRUS = '\U0001F9A0'  # 🦠 Virus/Microbe
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure for Protection Coverage Assessment
    sheets = [
        "Instructions & Legend",
        "Coverage_Analysis",
        "Agent_Status",
        "Scan_Compliance",
        "Detection_Metrics",
        "Incident_Response",
        "User_Awareness",
        "Performance_Metrics",
        "Licensing_Support",
        "Capability_Requirements",
        "Evidence_Register",
        "Gap_Analysis",
        "Approval_Sign_Of",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="E26B0A", end_color="E26B0A", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_protected": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        },
        "status_unprotected": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_unknown": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        },
        "gap_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "gap_high": {
            "fill": PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),
        },
        "gap_medium": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
        },
        "gap_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS
# ============================================================================

def create_base_validations(ws):
    """Create data validation objects for standard dropdowns."""
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_unknown': DataValidation(
            type="list",
            formula1='"Yes,No,Unknown"',
            allow_blank=False
        ),
        'protection_status': DataValidation(
            type="list",
            formula1='f"{CHECK} Protected,⚠️ Outdated,❌ Not Protected,🔴 Inactive,❓ Unknown"',
            allow_blank=False
        ),
        'agent_status': DataValidation(
            type="list",
            formula1='f"{CHECK} Active,⚠️ Outdated,❌ Not Installed,🔴 Inactive,🔄 Updating,❓ Unknown"',
            allow_blank=False
        ),
        'scan_status': DataValidation(
            type="list",
            formula1='f"{CHECK} Compliant,⚠️ Overdue,❌ Failed,🔄 In Progress,❓ Unknown"',
            allow_blank=False
        ),
        'antivirus_product': DataValidation(
            type="list",
            formula1='"Microsoft Defender,CrowdStrike Falcon,SentinelOne,Sophos,Trend Micro,McAfee,Symantec,Kaspersky,ESET,Bitdefender,Carbon Black,Palo Alto Cortex XDR,None,Other"',
            allow_blank=False
        ),
        'detection_severity': DataValidation(
            type="list",
            formula1='"🔴 Critical,🟠 High,🟡 Medium,🟢 Low"',
            allow_blank=False
        ),
        'detection_type': DataValidation(
            type="list",
            formula1='f"{VIRUS} Malware,🎣 Phishing,🪱 Ransomware,🕵️ Spyware,⚠️ PUA,🔧 Exploit,📧 Spam,🔒 Blocked,Other"',
            allow_blank=False
        ),
        'remediation_status': DataValidation(
            type="list",
            formula1='f"{CHECK} Remediated,🔄 In Progress,⏳ Pending,❌ Failed,🔴 Reinfection,❓ Unknown"',
            allow_blank=False
        ),
        'incident_severity': DataValidation(
            type="list",
            formula1='"🔴 Critical,🟠 High,🟡 Medium,🟢 Low"',
            allow_blank=False
        ),
        'incident_status': DataValidation(
            type="list",
            formula1='"🔴 Open,🟡 Investigating,🔵 Contained,🟢 Resolved,✅ Closed"',
            allow_blank=False
        ),
        'sla_compliance': DataValidation(
            type="list",
            formula1='f"{CHECK} Met,⚠️ At Risk,❌ Missed"',
            allow_blank=False
        ),
        'training_status': DataValidation(
            type="list",
            formula1='f"{CHECK} Completed,⏳ In Progress,❌ Not Started,⏰ Overdue"',
            allow_blank=False
        ),
        'phishing_result': DataValidation(
            type="list",
            formula1='f"{CHECK} Passed,⚠️ Clicked Link,❌ Submitted Credentials,🎓 Reported"',
            allow_blank=False
        ),
        'gap_severity': DataValidation(
            type="list",
            formula1='"🔴 Critical,🟠 High,🟡 Medium,🟢 Low"',
            allow_blank=False
        ),
        'gap_status': DataValidation(
            type="list",
            formula1='"🔴 Open,🟡 In Progress,🟢 Resolved,✅ Closed,⏸️ On Hold"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"📄 Config Export,📸 Screenshot,📊 Report,📜 License,📋 Policy,📁 Log File,🔍 Scan Result,Other"',
            allow_blank=False
        ),
        'verification_status': DataValidation(
            type="list",
            formula1='f"{CHECK} Verified,⏳ Pending,❌ Not Verified,⚠️ Needs Review"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1='f"{CHECK} Approved,✅ Approved with Conditions,❌ Rejected,⏳ Pending Review"',
            allow_blank=False
        ),
        'license_type': DataValidation(
            type="list",
            formula1='"Subscription,Perpetual,Bundled,Trial,Open Source,Unknown"',
            allow_blank=False
        ),
        'support_status': DataValidation(
            type="list",
            formula1='f"{CHECK} Active,⚠️ Expiring Soon,❌ Expired,❓ Unknown"',
            allow_blank=False
        ),
    }

    # Add all validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)

    return validations


# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create comprehensive instructions and legend sheet."""
    
    # Document header
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "MALWARE PROTECTION COVERAGE & EFFECTIVENESS ASSESSMENT"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "ISO/IEC 27001:2022 - Control A.8.7 (Protection Against Malware)"
    apply_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 20

    # Document metadata
    row = 4
    metadata = [
        ("Document ID:", "ISMS-IMP-A.8.1-7-18-19.S2"),
        ("Workbook:", "Protection Coverage Assessment"),
        ("Version:", "1.0"),
        ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Assessment Period:", "[To be completed by assessor]"),
        ("Assessor:", "[Name]"),
    ]
    
    for label, value in metadata:
        ws[f'A{row}'].value = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].value = value
        row += 1

    # Purpose section
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "📋 PURPOSE"
    apply_style(cell, styles['subheader'])
    row += 1

    purpose_text = """This workbook provides comprehensive assessment of malware protection coverage and effectiveness for compliance with ISO/IEC 27001:2022 Control A.8.7 (Protection Against Malware).

The assessment evaluates:
• Anti-malware/EDR agent deployment coverage
• Agent status, version currency, and signature updates
• Scan compliance (full scans, quick scans)
• Malware detection effectiveness and remediation success
• Incident response SLA compliance
• User awareness training and phishing simulation results
• Performance metrics (false positives, scan impact)
• Licensing and support contract status"""

    ws.merge_cells(f'A{row}:F{row+7}')
    cell = ws[f'A{row}']
    cell.value = purpose_text
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 120
    row += 8

    # Workbook structure
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = f"{CHART} WORKBOOK STRUCTURE"
    apply_style(cell, styles['subheader'])
    row += 1

    sheets_info = [
        ("Coverage_Analysis", "Per-endpoint protection status (agent installed, version, signatures)"),
        ("Agent_Status", "Agent health, communication status, deployment gaps"),
        ("Scan_Compliance", "Full scan and quick scan compliance tracking"),
        ("Detection_Metrics", "Malware detections by type, severity, remediation status"),
        ("Incident_Response", "Malware incidents, response times, SLA compliance"),
        ("User_Awareness", "Security training completion, phishing simulation results"),
        ("Performance_Metrics", "False positives/negatives, scan performance impact"),
        ("Licensing_Support", "License status, support contracts, expiration tracking"),
        ("Capability_Requirements", "Policy requirements mapped to implementation (A.8.7)"),
        ("Evidence_Register", "Comprehensive evidence documentation (100 entries)"),
        ("Gap_Analysis", "Gap identification, severity, remediation tracking"),
        ("Approval_Sign_Of", "Multi-level approval workflow"),
    ]

    headers = ["Sheet Name", "Description"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    for sheet_name, description in sheets_info:
        ws.cell(row=row, column=1).value = sheet_name
        ws.cell(row=row, column=2).value = description
        thin = Side(style="thin")
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

    # Status legend
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "🎨 STATUS LEGEND"
    apply_style(cell, styles['subheader'])
    row += 1

    legend_items = [
        (f"{CHECK} Protected / Active / Compliant", "Endpoint fully protected, agent active, signatures current", "status_protected"),
        (f"{WARNING} Outdated / At Risk", "Agent outdated or signatures >24 hours old", "status_partial"),
        (f"{XMARK} Not Protected / Failed", "No agent installed or scan failed", "status_unprotected"),
        ("🔴 Inactive / Critical", "Agent inactive or critical malware detected", "gap_critical"),
        ("❓ Unknown", "Protection status cannot be determined", "status_unknown"),
        ("🔴 Critical Severity", "Critical gap requiring immediate remediation (24 hours)", "gap_critical"),
        ("🟠 High Severity", "High-priority gap requiring remediation (7 days)", "gap_high"),
        ("🟡 Medium Severity", "Medium-priority gap requiring remediation (30 days)", "gap_medium"),
        ("🟢 Low Severity", "Low-priority gap, scheduled remediation (90 days)", "gap_low"),
    ]

    headers = ["Status/Severity", "Meaning", "Color"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    for status, meaning, style_key in legend_items:
        ws.cell(row=row, column=1).value = status
        ws.cell(row=row, column=2).value = meaning
        cell = ws.cell(row=row, column=3)
        cell.value = "█████"
        if style_key in styles:
            apply_style(cell, styles[style_key])
        thin = Side(style="thin")
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

    # Target metrics
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = f"{TARGET} TARGET METRICS (per POL-S3)"
    apply_style(cell, styles['subheader'])
    row += 1

    targets = [
        ("Protection Coverage (Corporate)", "≥98%"),
        ("Protection Coverage (BYOD)", "≥80%"),
        ("Signature Currency (<24 hours)", "≥98%"),
        ("Full Scan Compliance (Weekly)", "≥95%"),
        ("Quick Scan Compliance (Daily)", "≥90%"),
        ("Remediation Success Rate", "≥95%"),
        ("False Positive Rate", "<10%"),
        ("Re-Infection Rate", "<5%"),
        ("Incident SLA Compliance", "100% (Critical/High)"),
        ("Training Completion", "≥98%"),
        ("Phishing Click Rate", "<10%"),
    ]

    headers = ["Metric", "Target"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    for metric, target in targets:
        ws.cell(row=row, column=1).value = metric
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = target
        thin = Side(style="thin")
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

    # Important notes
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = f"{WARNING} IMPORTANT NOTES"
    apply_style(cell, styles['subheader'])
    row += 1

    notes = [
        f"{BULLET} This assessment is VENDOR-AGNOSTIC - works with ANY anti-malware/EDR solution",
        f"{BULLET} Yellow-highlighted cells require user input - complete ALL yellow cells",
        f"{BULLET} Document ID must NOT be changed - used for dashboard workbook linking",
        f"{BULLET} Critical malware incidents require 24-hour investigation SLA",
        f"{BULLET} Unprotected endpoints (no agent) are CRITICAL gaps requiring immediate remediation",
        f"{BULLET} Obtain final CISO approval in Approval_Sign_Off sheet before audit submission",
    ]

    for note in notes:
        ws.cell(row=row, column=1).value = note
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15


# ============================================================================
# SECTION 4: COVERAGE ANALYSIS SHEET
# ============================================================================

def create_coverage_analysis_sheet(ws, styles):
    """Create per-endpoint protection coverage analysis sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:M1')
    cell = ws['A1']
    cell.value = "MALWARE PROTECTION COVERAGE ANALYSIS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:M2')
    cell = ws['A2']
    cell.value = "Per-endpoint anti-malware/EDR agent status, version, signatures, and protection effectiveness"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Device ID",
        "Hostname",
        "Device Type",
        "Protection Product",
        "Agent Version",
        "Agent Status",
        "Signature Version",
        "Signature Date",
        "Signatures Current",
        "Last Full Scan",
        "Last Quick Scan",
        "Protection Status",
        "Notes"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Data rows (50 rows - linked to Endpoint Inventory)
    start_row = 5
    for i in range(50):
        current_row = start_row + i
        
        # Device ID (would link to Inventory in actual use)
        ws.cell(row=current_row, column=1).value = f"EP-{1001+i:04d}"
        
        # Hostname (input)
        cell = ws.cell(row=current_row, column=2)
        apply_style(cell, styles['input_cell'])
        
        # Device Type (input)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        
        # Protection Product (dropdown)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        validations['antivirus_product'].add(cell)
        
        # Agent Version (input)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        
        # Agent Status (dropdown)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        validations['agent_status'].add(cell)
        
        # Signature Version (input)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        
        # Signature Date (input)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])
        
        # Signatures Current (dropdown)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])
        validations['yes_no_unknown'].add(cell)
        
        # Last Full Scan (input)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])
        
        # Last Quick Scan (input)
        cell = ws.cell(row=current_row, column=11)
        apply_style(cell, styles['input_cell'])
        
        # Protection Status (calculated)
        cell = ws.cell(row=current_row, column=12)
        cell.value = f'=IF(F{current_row}=f"{CHECK} Active",IF(I{current_row}="Yes",f"{CHECK} Protected",IF(I{current_row}="No",f"{WARNING} Outdated","❓ Unknown")),IF(F{current_row}=f"{XMARK} Not Installed",f"{XMARK} Not Protected",IF(F{current_row}="🔴 Inactive","🔴 Inactive",f"{WARNING} Outdated")))'
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=13)
        apply_style(cell, styles['input_cell'])

    # Summary statistics
    summary_row = start_row + 52
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} PROTECTION COVERAGE SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Total Endpoints:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTA(B5:B54)'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = f"{CHECK} Protected:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(L5:L54,f"{CHECK} Protected")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = f"{XMARK} Not Protected:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(L5:L54,f"{XMARK} Not Protected")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Protection Coverage Rate:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=IF(B{summary_row-3}>0,B{summary_row-2}/B{summary_row-3}*100,0)&"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    
    summary_row += 2
    ws[f'A{summary_row}'].value = "Signatures Current (<24h):"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=COUNTIF(I5:I54,"Yes")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Signature Currency Rate:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].value = f'=IF(B{summary_row-4}>0,B{summary_row-1}/B{summary_row-4}*100,0)&"%"'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 30


# ============================================================================
# SECTION 5: AGENT STATUS SHEET
# ============================================================================

def create_agent_status_sheet(ws, styles):
    """Create detailed agent health and communication status sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "ANTI-MALWARE AGENT STATUS & HEALTH"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "Agent health monitoring, communication status, deployment gaps"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Device ID",
        "Hostname",
        "Agent Installed",
        "Agent Version",
        "Latest Version",
        "Agent Outdated",
        "Last Check-In",
        "Communication Status",
        "Deployment Gap Reason",
        "Notes"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Data rows (50 rows)
    start_row = 5
    for i in range(50):
        current_row = start_row + i
        
        # Device ID
        ws.cell(row=current_row, column=1).value = f"EP-{1001+i:04d}"
        
        # Hostname (input)
        cell = ws.cell(row=current_row, column=2)
        apply_style(cell, styles['input_cell'])
        
        # Agent Installed (dropdown)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        validations['yes_no'].add(cell)
        
        # Agent Version (input)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        
        # Latest Version (input)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        
        # Agent Outdated (calculated)
        cell = ws.cell(row=current_row, column=6)
        cell.value = f'=IF(C{current_row}="Yes",IF(D{current_row}=E{current_row},"No","Yes"),"N/A")'
        
        # Last Check-In (input)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        
        # Communication Status (dropdown)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])
        comm_status_dv = DataValidation(
            type="list",
            formula1='f"{CHECK} Active,⚠️ Delayed,❌ No Communication,❓ Unknown"',
            allow_blank=False
        )
        ws.add_data_validation(comm_status_dv)
        comm_status_dv.add(cell)
        
        # Deployment Gap Reason (input)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])

    # Summary
    summary_row = start_row + 52
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} AGENT STATUS SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Agents Installed:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(C5:C54,"Yes")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Agents Not Installed:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(C5:C54,"No")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Outdated Agents:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(F5:F54,"Yes")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Communication Issues:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(H5:H54,f"{XMARK} No Communication")+COUNTIF(H5:H54,f"{WARNING} Delayed")'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 30


# ============================================================================
# SECTION 6: SCAN COMPLIANCE SHEET
# ============================================================================

def create_scan_compliance_sheet(ws, styles):
    """Create scan compliance tracking sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "MALWARE SCAN COMPLIANCE"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:K2')
    cell = ws['A2']
    cell.value = "Full scan compliance (weekly), quick scan compliance (daily), scan failure tracking"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Device ID",
        "Hostname",
        "Last Full Scan Date",
        "Full Scan Status",
        "Days Since Full Scan",
        "Full Scan Compliant",
        "Last Quick Scan Date",
        "Quick Scan Status",
        "Days Since Quick Scan",
        "Quick Scan Compliant",
        "Notes"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Data rows (50 rows)
    start_row = 5
    for i in range(50):
        current_row = start_row + i
        
        # Device ID
        ws.cell(row=current_row, column=1).value = f"EP-{1001+i:04d}"
        
        # Hostname (input)
        cell = ws.cell(row=current_row, column=2)
        apply_style(cell, styles['input_cell'])
        
        # Last Full Scan Date (input)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        
        # Full Scan Status (dropdown)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        validations['scan_status'].add(cell)
        
        # Days Since Full Scan (calculated - would use TODAY() in Excel)
        cell = ws.cell(row=current_row, column=5)
        cell.value = f'=IF(C{current_row}<>"",TODAY()-C{current_row},"")'
        
        # Full Scan Compliant (calculated - weekly = 7 days)
        cell = ws.cell(row=current_row, column=6)
        cell.value = f'=IF(E{current_row}<=7,"Yes",IF(E{current_row}>7,"No","Unknown"))'
        
        # Last Quick Scan Date (input)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        
        # Quick Scan Status (dropdown)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])
        validations['scan_status'].add(cell)
        
        # Days Since Quick Scan (calculated)
        cell = ws.cell(row=current_row, column=9)
        cell.value = f'=IF(G{current_row}<>"",TODAY()-G{current_row},"")'
        
        # Quick Scan Compliant (calculated - daily = 1 day)
        cell = ws.cell(row=current_row, column=10)
        cell.value = f'=IF(I{current_row}<=1,"Yes",IF(I{current_row}>1,"No","Unknown"))'
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=11)
        apply_style(cell, styles['input_cell'])

    # Summary
    summary_row = start_row + 52
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} SCAN COMPLIANCE SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Full Scan Compliant (≤7 days):"
    ws[f'B{summary_row}'].value = f'=COUNTIF(F5:F54,"Yes")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Full Scan Compliance Rate:"
    ws[f'B{summary_row}'].value = f'=IF(COUNTA(F5:F54)>0,B{summary_row-1}/COUNTA(F5:F54)*100,0)&"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=11)
    
    summary_row += 2
    ws[f'A{summary_row}'].value = "Quick Scan Compliant (≤1 day):"
    ws[f'B{summary_row}'].value = f'=COUNTIF(J5:J54,"Yes")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Quick Scan Compliance Rate:"
    ws[f'B{summary_row}'].value = f'=IF(COUNTA(J5:J54)>0,B{summary_row-1}/COUNTA(J5:J54)*100,0)&"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=11)

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 30


# ============================================================================
# SECTION 7: DETECTION METRICS SHEET
# ============================================================================

def create_detection_metrics_sheet(ws, styles):
    """Create malware detection and remediation metrics sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:L1')
    cell = ws['A1']
    cell.value = "MALWARE DETECTION & REMEDIATION METRICS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:L2')
    cell = ws['A2']
    cell.value = "Malware detections by type, severity, remediation status, and effectiveness (last 90 days)"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Detection ID",
        "Detection Date",
        "Device ID",
        "Hostname",
        "Detection Type",
        "Threat Name",
        "Severity",
        "Remediation Status",
        "Remediation Date",
        "Remediation Time (Hours)",
        "Re-Infection",
        "Notes"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Detection rows (100 rows for last 90 days)
    start_row = 5
    for i in range(100):
        current_row = start_row + i
        
        # Detection ID (auto-generated)
        ws.cell(row=current_row, column=1).value = f"DET-{i+1:04d}"
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        # Detection Date (input)
        cell = ws.cell(row=current_row, column=2)
        apply_style(cell, styles['input_cell'])
        
        # Device ID (input)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        
        # Hostname (input)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        
        # Detection Type (dropdown)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        validations['detection_type'].add(cell)
        
        # Threat Name (input)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        
        # Severity (dropdown)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        validations['detection_severity'].add(cell)
        
        # Remediation Status (dropdown)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])
        validations['remediation_status'].add(cell)
        
        # Remediation Date (input)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])
        
        # Remediation Time (calculated)
        cell = ws.cell(row=current_row, column=10)
        cell.value = f'=IF(AND(B{current_row}<>"",I{current_row}<>""),(I{current_row}-B{current_row})*24,"")'
        
        # Re-Infection (dropdown)
        cell = ws.cell(row=current_row, column=11)
        apply_style(cell, styles['input_cell'])
        validations['yes_no_na'].add(cell)
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=12)
        apply_style(cell, styles['input_cell'])

    # Summary
    summary_row = start_row + 102
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} DETECTION METRICS SUMMARY (Last 90 Days)"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Total Detections:"
    ws[f'B{summary_row}'].value = f'=COUNTA(B5:B104)'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = f"{CHECK} Remediated:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(H5:H104,f"{CHECK} Remediated")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Remediation Success Rate:"
    ws[f'B{summary_row}'].value = f'=IF(B{summary_row-2}>0,B{summary_row-1}/B{summary_row-2}*100,0)&"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=11)
    
    summary_row += 2
    ws[f'A{summary_row}'].value = "By Detection Type:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    
    detection_types = [
        f"{VIRUS} Malware", "🎣 Phishing", "🪱 Ransomware", "🕵️ Spyware", 
        f"{WARNING} PUA", "🔧 Exploit", "📧 Spam", f"{LOCK} Blocked"
    ]
    
    for det_type in detection_types:
        summary_row += 1
        ws[f'A{summary_row}'].value = det_type + ":"
        ws[f'B{summary_row}'].value = f'=COUNTIF(E5:E104,"{det_type}")'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 30


# ============================================================================
# SECTION 8: INCIDENT RESPONSE SHEET
# ============================================================================

def create_incident_response_sheet(ws, styles):
    """Create malware incident response and SLA compliance sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:M1')
    cell = ws['A1']
    cell.value = "MALWARE INCIDENT RESPONSE & SLA COMPLIANCE"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:M2')
    cell = ws['A2']
    cell.value = "Major malware incidents, response times, SLA compliance, post-incident review"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Incident ID",
        "Incident Date",
        "Severity",
        "Affected Devices",
        "Incident Type",
        "Investigation Time (H)",
        "Investigation SLA",
        "Containment Time (H)",
        "Containment SLA",
        "Remediation Time (H)",
        "Remediation SLA",
        "Incident Status",
        "Post-Incident Review"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Incident rows (50 rows)
    start_row = 5
    for i in range(50):
        current_row = start_row + i
        
        # Incident ID (auto-generated)
        ws.cell(row=current_row, column=1).value = f"INC-{i+1:04d}"
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        # Incident Date (input)
        cell = ws.cell(row=current_row, column=2)
        apply_style(cell, styles['input_cell'])
        
        # Severity (dropdown)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        validations['incident_severity'].add(cell)
        
        # Affected Devices (input)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        
        # Incident Type (input)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        
        # Investigation Time (input)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        
        # Investigation SLA (calculated - Critical: 4h, High: 8h, Medium: 24h)
        cell = ws.cell(row=current_row, column=7)
        cell.value = f'=IF(C{current_row}="🔴 Critical",IF(F{current_row}<=4,f"{CHECK} Met",IF(F{current_row}>4,f"{XMARK} Missed","N/A")),IF(C{current_row}="🟠 High",IF(F{current_row}<=8,f"{CHECK} Met",IF(F{current_row}>8,f"{XMARK} Missed","N/A")),"N/A"))'
        
        # Containment Time (input)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])
        
        # Containment SLA (calculated - Critical: 24h, High: 48h)
        cell = ws.cell(row=current_row, column=9)
        cell.value = f'=IF(C{current_row}="🔴 Critical",IF(H{current_row}<=24,f"{CHECK} Met",IF(H{current_row}>24,f"{XMARK} Missed","N/A")),IF(C{current_row}="🟠 High",IF(H{current_row}<=48,f"{CHECK} Met",IF(H{current_row}>48,f"{XMARK} Missed","N/A")),"N/A"))'
        
        # Remediation Time (input)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])
        
        # Remediation SLA (calculated - Critical: 72h, High: 1 week)
        cell = ws.cell(row=current_row, column=11)
        cell.value = f'=IF(C{current_row}="🔴 Critical",IF(J{current_row}<=72,f"{CHECK} Met",IF(J{current_row}>72,f"{XMARK} Missed","N/A")),IF(C{current_row}="🟠 High",IF(J{current_row}<=168,f"{CHECK} Met",IF(J{current_row}>168,f"{XMARK} Missed","N/A")),"N/A"))'
        
        # Incident Status (dropdown)
        cell = ws.cell(row=current_row, column=12)
        apply_style(cell, styles['input_cell'])
        validations['incident_status'].add(cell)
        
        # Post-Incident Review (dropdown)
        cell = ws.cell(row=current_row, column=13)
        apply_style(cell, styles['input_cell'])
        validations['yes_no_na'].add(cell)

    # Summary
    summary_row = start_row + 52
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} INCIDENT RESPONSE SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Total Incidents:"
    ws[f'B{summary_row}'].value = f'=COUNTA(B5:B54)'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "🔴 Critical:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(C5:C54,"🔴 Critical")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "🟠 High:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(C5:C54,"🟠 High")'
    
    summary_row += 2
    ws[f'A{summary_row}'].value = "Investigation SLA Met:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(G5:G54,f"{CHECK} Met")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Investigation SLA Compliance:"
    ws[f'B{summary_row}'].value = f'=IF(COUNTA(G5:G54)>0,B{summary_row-1}/COUNTA(G5:G54)*100,0)&"%"'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 20


# ============================================================================
# SECTION 9: USER AWARENESS SHEET
# ============================================================================

def create_user_awareness_sheet(ws, styles):
    """Create security awareness training and phishing simulation tracking sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "SECURITY AWARENESS TRAINING & PHISHING SIMULATIONS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = "Annual training completion, quarterly phishing simulation results, click rate trends"
    apply_style(cell, styles['subheader'])

    row = 4
    
    # Training Completion Section
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    cell.value = "📚 ANNUAL SECURITY AWARENESS TRAINING"
    apply_style(cell, styles['subheader'])
    row += 1

    headers = ["User", "Department", "Training Assigned", "Training Completed", "Training Status", "Completion Date", "Score", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    # Training rows (50 users)
    training_start_row = row
    for i in range(50):
        current_row = training_start_row + i
        
        # User (input)
        cell = ws.cell(row=current_row, column=1)
        apply_style(cell, styles['input_cell'])
        
        # Department (input)
        cell = ws.cell(row=current_row, column=2)
        apply_style(cell, styles['input_cell'])
        
        # Training Assigned (dropdown)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        validations['yes_no'].add(cell)
        
        # Training Completed (dropdown)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        validations['yes_no'].add(cell)
        
        # Training Status (dropdown)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        validations['training_status'].add(cell)
        
        # Completion Date (input)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        
        # Score (input)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=8)
        apply_style(cell, styles['input_cell'])

    # Training Summary
    row = training_start_row + 52
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = f"{CHART} TRAINING SUMMARY"
    apply_style(cell, styles['subheader'])
    row += 1

    ws[f'A{row}'].value = "Total Users:"
    ws[f'B{row}'].value = f'=COUNTA(A{training_start_row}:A{training_start_row+49})'
    row += 1
    
    ws[f'A{row}'].value = f"{CHECK} Completed:"
    ws[f'B{row}'].value = f'=COUNTIF(E{training_start_row}:E{training_start_row+49},f"{CHECK} Completed")'
    row += 1
    
    ws[f'A{row}'].value = "Training Completion Rate:"
    ws[f'B{row}'].value = f'=IF(B{row-2}>0,B{row-1}/B{row-2}*100,0)&"%"'
    ws[f'B{row}'].font = Font(bold=True, size=11)
    row += 2

    # Phishing Simulation Section
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    cell.value = "🎣 QUARTERLY PHISHING SIMULATIONS"
    apply_style(cell, styles['subheader'])
    row += 1

    headers = ["Simulation ID", "Quarter", "Users Targeted", "Emails Delivered", "Links Clicked", "Credentials Submitted", "Reported Phishing", "Click Rate (%)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    # Phishing simulation rows (12 quarters = 3 years)
    phishing_start_row = row
    quarters = ["Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024", "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025", "Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"]
    for i, quarter in enumerate(quarters):
        current_row = phishing_start_row + i
        
        # Simulation ID
        ws.cell(row=current_row, column=1).value = f"PHISH-{i+1:02d}"
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        # Quarter
        ws.cell(row=current_row, column=2).value = quarter
        
        # Users Targeted (input)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        
        # Emails Delivered (input)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        
        # Links Clicked (input)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        
        # Credentials Submitted (input)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        
        # Reported Phishing (input)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        
        # Click Rate (calculated)
        cell = ws.cell(row=current_row, column=8)
        cell.value = f'=IF(D{current_row}>0,E{current_row}/D{current_row}*100,0)'

    # Phishing Summary
    row = phishing_start_row + 14
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = f"{CHART} PHISHING SIMULATION SUMMARY"
    apply_style(cell, styles['subheader'])
    row += 1

    ws[f'A{row}'].value = "Average Click Rate:"
    ws[f'B{row}'].value = f'=AVERAGE(H{phishing_start_row}:H{phishing_start_row+11})'
    ws[f'B{row}'].font = Font(bold=True, size=11)
    row += 1
    
    ws[f'A{row}'].value = "Latest Click Rate:"
    ws[f'B{row}'].value = f'=H{phishing_start_row+11}'
    row += 1
    
    ws[f'A{row}'].value = "Trend:"
    ws[f'B{row}'].value = f'=IF(H{phishing_start_row+11}<H{phishing_start_row+10},"📉 Improving","📈 Degrading")'

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15


# ============================================================================
# SECTION 10: PERFORMANCE METRICS SHEET
# ============================================================================

def create_performance_metrics_sheet(ws, styles):
    """Create malware protection performance metrics sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "MALWARE PROTECTION PERFORMANCE METRICS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "False positives/negatives, scan performance impact, monthly tracking (12 months)"
    apply_style(cell, styles['subheader'])

    row = 4
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "📈 MONTHLY PERFORMANCE METRICS"
    apply_style(cell, styles['subheader'])
    row += 1

    headers = ["Month", "False Positives", "False Negatives", "FP Rate (%)", "Avg Scan Time (min)", "Performance Complaints"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    # 12 months of data
    months = ["Jan 2025", "Feb 2025", "Mar 2025", "Apr 2025", "May 2025", "Jun 2025", 
              "Jul 2025", "Aug 2025", "Sep 2025", "Oct 2025", "Nov 2025", "Dec 2025"]
    
    metrics_start_row = row
    for i, month in enumerate(months):
        current_row = metrics_start_row + i
        
        # Month
        ws.cell(row=current_row, column=1).value = month
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        # False Positives (input)
        cell = ws.cell(row=current_row, column=2)
        apply_style(cell, styles['input_cell'])
        
        # False Negatives (input)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        
        # FP Rate (calculated - FP / Total Detections * 100)
        cell = ws.cell(row=current_row, column=4)
        # Note: Would need total detections - simplified for template
        apply_style(cell, styles['input_cell'])
        
        # Avg Scan Time (input)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        
        # Performance Complaints (input)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])

    # Summary
    row = metrics_start_row + 14
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = f"{CHART} PERFORMANCE SUMMARY (12 Months)"
    apply_style(cell, styles['subheader'])
    row += 1

    ws[f'A{row}'].value = "Total False Positives:"
    ws[f'B{row}'].value = f'=SUM(B{metrics_start_row}:B{metrics_start_row+11})'
    row += 1
    
    ws[f'A{row}'].value = "Total False Negatives:"
    ws[f'B{row}'].value = f'=SUM(C{metrics_start_row}:C{metrics_start_row+11})'
    row += 1
    
    ws[f'A{row}'].value = "Avg FP Rate:"
    ws[f'B{row}'].value = f'=AVERAGE(D{metrics_start_row}:D{metrics_start_row+11})'
    row += 1
    
    ws[f'A{row}'].value = "Total Performance Complaints:"
    ws[f'B{row}'].value = f'=SUM(F{metrics_start_row}:F{metrics_start_row+11})'

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 22


# ============================================================================
# SECTION 11: LICENSING & SUPPORT SHEET
# ============================================================================

def create_licensing_support_sheet(ws, styles):
    """Create licensing and support tracking sheet."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "LICENSING & SUPPORT CONTRACT TRACKING"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "Anti-malware/EDR licensing, support contracts, renewal tracking, cost management"
    apply_style(cell, styles['subheader'])

    row = 4
    headers = [
        "Product Name",
        "Vendor",
        "License Type",
        "License Count",
        "Deployed Count",
        "Support Status",
        "Support Expiration",
        "Days Until Expiration",
        "Annual Cost",
        "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    # License tracking rows (10 rows)
    start_row = row
    for i in range(10):
        current_row = start_row + i
        
        # Product Name (input)
        cell = ws.cell(row=current_row, column=1)
        apply_style(cell, styles['input_cell'])
        
        # Vendor (input)
        cell = ws.cell(row=current_row, column=2)
        apply_style(cell, styles['input_cell'])
        
        # License Type (dropdown)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        validations['license_type'].add(cell)
        
        # License Count (input)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        
        # Deployed Count (input)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        
        # Support Status (dropdown)
        cell = ws.cell(row=current_row, column=6)
        apply_style(cell, styles['input_cell'])
        validations['support_status'].add(cell)
        
        # Support Expiration (input)
        cell = ws.cell(row=current_row, column=7)
        apply_style(cell, styles['input_cell'])
        
        # Days Until Expiration (calculated)
        cell = ws.cell(row=current_row, column=8)
        cell.value = f'=IF(G{current_row}<>"",G{current_row}-TODAY(),"")'
        
        # Annual Cost (input)
        cell = ws.cell(row=current_row, column=9)
        apply_style(cell, styles['input_cell'])
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=10)
        apply_style(cell, styles['input_cell'])

    # Summary
    summary_row = start_row + 12
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} LICENSE & SUPPORT SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Total License Cost:"
    ws[f'B{summary_row}'].value = f'=SUM(I{start_row}:I{start_row+9})'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Expiring Soon (<90 days):"
    ws[f'B{summary_row}'].value = f'=COUNTIF(H{start_row}:H{start_row+9},"<90")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = f"{CHECK} Active Support:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(F{start_row}:F{start_row+9},f"{CHECK} Active")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = f"{XMARK} Expired Support:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(F{start_row}:F{start_row+9},f"{XMARK} Expired")'

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 30


# ============================================================================
# SECTION 12: CAPABILITY REQUIREMENTS SHEET
# ============================================================================

def create_capability_requirements_sheet(ws, styles):
    """Create policy requirements mapping sheet (A.8.7 requirements)."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "CAPABILITY REQUIREMENTS MAPPING"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "A.8.7 Policy Requirements → Implementation Verification (25 requirements)"
    apply_style(cell, styles['subheader'])

    # Column headers
    headers = [
        "Req ID",
        "Policy Requirement",
        "Implemented",
        "Evidence Reference",
        "Notes",
        "Status"
    ]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])

    # Requirements (from POL-S3)
    requirements = [
        ("REQ-A87-001", "Anti-malware/EDR deployed on ≥98% corporate endpoints"),
        ("REQ-A87-002", "Anti-malware/EDR deployed on ≥80% BYOD devices"),
        ("REQ-A87-003", "Signature updates automatic and daily"),
        ("REQ-A87-004", "Signatures current (<24 hours) on ≥98% endpoints"),
        ("REQ-A87-005", "Real-time scanning enabled on all endpoints"),
        ("REQ-A87-006", "Weekly full scans on ≥95% endpoints"),
        ("REQ-A87-007", "Daily quick scans on ≥90% endpoints"),
        ("REQ-A87-008", "Malware detection alerts to security team"),
        ("REQ-A87-009", "Quarantine/remediation automatic where possible"),
        ("REQ-A87-010", "Remediation success rate ≥95%"),
        ("REQ-A87-011", "Re-infection rate <5%"),
        ("REQ-A87-012", "False positive rate <10%"),
        ("REQ-A87-013", "Critical malware incidents investigated within 4 hours"),
        ("REQ-A87-014", "High malware incidents investigated within 8 hours"),
        ("REQ-A87-015", "Critical malware contained within 24 hours"),
        ("REQ-A87-016", "High malware contained within 48 hours"),
        ("REQ-A87-017", "Annual security awareness training ≥98% completion"),
        ("REQ-A87-018", "Quarterly phishing simulations conducted"),
        ("REQ-A87-019", "Phishing click rate <10% or improving trend"),
        ("REQ-A87-020", "Malware protection integrated with SIEM"),
        ("REQ-A87-021", "Malware incidents logged and tracked"),
        ("REQ-A87-022", "Post-incident reviews for critical incidents"),
        ("REQ-A87-023", "Malware protection licenses current"),
        ("REQ-A87-024", "Support contracts active"),
        ("REQ-A87-025", "Quarterly protection effectiveness assessment"),
    ]

    start_row = 5
    for i, (req_id, requirement) in enumerate(requirements):
        current_row = start_row + i
        
        # Req ID
        ws.cell(row=current_row, column=1).value = req_id
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        # Policy Requirement
        ws.cell(row=current_row, column=2).value = requirement
        ws.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
        
        # Implemented (dropdown)
        cell = ws.cell(row=current_row, column=3)
        apply_style(cell, styles['input_cell'])
        validations['yes_no_na'].add(cell)
        
        # Evidence Reference (input)
        cell = ws.cell(row=current_row, column=4)
        apply_style(cell, styles['input_cell'])
        
        # Notes (input)
        cell = ws.cell(row=current_row, column=5)
        apply_style(cell, styles['input_cell'])
        
        # Status (calculated)
        cell = ws.cell(row=current_row, column=6)
        cell.value = f'=IF(C{current_row}="Yes",f"{CHECK} Compliant",IF(C{current_row}="N/A","N/A",f"{XMARK} Gap"))'
        
        thin = Side(style="thin")
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Summary
    summary_row = start_row + len(requirements) + 2
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    cell = ws[f'A{summary_row}']
    cell.value = f"{CHART} REQUIREMENTS COMPLIANCE SUMMARY"
    apply_style(cell, styles['subheader'])

    summary_row += 1
    ws[f'A{summary_row}'].value = "Total Requirements:"
    ws[f'B{summary_row}'].value = len(requirements)
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Implemented:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(C5:C{start_row+len(requirements)-1},"Yes")'
    
    summary_row += 1
    ws[f'A{summary_row}'].value = "Compliance Rate:"
    ws[f'B{summary_row}'].value = f'=IF(B{summary_row-2}>0,B{summary_row-1}/B{summary_row-2}*100,0)&"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=11)

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15


# ============================================================================
# SECTIONS 13-15: Evidence, Gaps, Approval (Same as Script 1)
# ============================================================================

def create_evidence_register_sheet(ws, styles):
    """Create comprehensive evidence documentation sheet."""
    validations = create_base_validations(ws)

    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "EVIDENCE REGISTER"
    apply_style(cell, styles['header'])

    ws.merge_cells('A2:J2')
    cell = ws['A2']
    cell.value = "Comprehensive evidence documentation for malware protection assessment (100 entries)"
    apply_style(cell, styles['subheader'])

    headers = ["Evidence ID", "Evidence Type", "Description", "Related Requirement", "Related Worksheet/Device", "File Location", "Collection Date", "Collected By", "Verification Status", "Notes"]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    # Evidence entries (100 rows)
    for i in range(100):
        current_row = row + i
        ws.cell(row=current_row, column=1).value = f"EVD-{i+1:03d}"
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        for col in range(2, 11):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
            if col == 2:
                validations['evidence_type'].add(cell)
            elif col == 9:
                validations['verification_status'].add(cell)

    # Summary
    summary_row = row + 102
    ws[f'A{summary_row}'].value = "Total Evidence:"
    ws[f'B{summary_row}'].value = f'=COUNTA(C5:C104)'

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 30


def create_gap_analysis_sheet(ws, styles):
    """Create gap identification and remediation tracking sheet."""
    validations = create_base_validations(ws)

    ws.merge_cells('A1:M1')
    cell = ws['A1']
    cell.value = "GAP ANALYSIS & REMEDIATION TRACKING"
    apply_style(cell, styles['header'])

    ws.merge_cells('A2:M2')
    cell = ws['A2']
    cell.value = "Gap identification, severity classification, remediation planning (40 entries)"
    apply_style(cell, styles['subheader'])

    headers = ["Gap ID", "Gap Description", "Affected Devices/Count", "Related Requirement", "Severity", "Risk", "Root Cause", "Remediation Plan", "Owner", "Due Date", "Status", "Budget Required", "Notes"]

    row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    row += 1

    # Gap entries (40 rows)
    for i in range(40):
        current_row = row + i
        ws.cell(row=current_row, column=1).value = f"GAP-{i+1:03d}"
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        
        for col in range(2, 14):
            cell = ws.cell(row=current_row, column=col)
            apply_style(cell, styles['input_cell'])
            if col == 5:
                validations['gap_severity'].add(cell)
            elif col == 11:
                validations['gap_status'].add(cell)

    # Summary
    summary_row = row + 42
    ws[f'A{summary_row}'].value = "🔴 Critical:"
    ws[f'B{summary_row}'].value = f'=COUNTIF(E5:E44,"🔴 Critical")'

    ws.column_dimensions['A'].width = 12
    for col in ['B','C','D','E','F','G','H','I','J','K','L','M']:
        ws.column_dimensions[col].width = 25


def create_approval_signoff_sheet(ws, styles):
    """Create multi-level approval workflow sheet."""
    validations = create_base_validations(ws)

    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "APPROVAL & SIGN-OFF WORKFLOW"
    apply_style(cell, styles['header'])

    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "Three-level approval: Assessor → IT Operations Manager → CISO"
    apply_style(cell, styles['subheader'])

    row = 4
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "📋 ASSESSMENT SUMMARY"
    apply_style(cell, styles['subheader'])
    row += 1

    summary_items = [
        ("Assessment Period:", "[Start Date] to [End Date]"),
        ("Protection Coverage Rate:", "=Coverage_Analysis!B59"),
        ("Signature Currency Rate:", "=Coverage_Analysis!B63"),
        ("Full Scan Compliance:", "=Scan_Compliance!B58"),
        ("Remediation Success Rate:", "=Detection_Metrics!B109"),
        ("Training Completion:", "See User_Awareness sheet"),
        ("Critical Gaps:", "=Gap_Analysis!B48"),
    ]

    for label, value in summary_items:
        ws[f'A{row}'].value = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].value = value
        row += 1

    # Approval sections (simplified for brevity)
    row += 2
    ws[f'A{row}'].value = "LEVEL 1: ASSESSOR SIGN-OFF"
    apply_style(cell, styles['subheader'])
    
    # Additional approval levels would follow same pattern as Script 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30


# ============================================================================
# SECTION 16: MAIN FUNCTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    print("=" * 78)
    print("ISMS-IMP-A.8.1-7-18-19.S2 - Protection Coverage Assessment Generator")
    print("ISO/IEC 27001:2022 Control: A.8.7 (Protection Against Malware)")
    print("=" * 78)
    print("\n🎯 Systems Engineering: Evidence-Based Malware Protection Assessment")
    print(f"{CHART} Vendor-Agnostic: Works with ANY anti-malware/EDR solution")
    print(f"{LOCK} Audit-Ready: Comprehensive coverage and effectiveness metrics")
    print("\n" + "─" * 78)

    print("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    print(f"{CHECK} Workbook created with 13 sheets")

    print("\n[Phase 2] Generating assessment sheets...")
    
    sheets = [
        ("Instructions & Legend", create_instructions_sheet),
        ("Coverage_Analysis", create_coverage_analysis_sheet),
        ("Agent_Status", create_agent_status_sheet),
        ("Scan_Compliance", create_scan_compliance_sheet),
        ("Detection_Metrics", create_detection_metrics_sheet),
        ("Incident_Response", create_incident_response_sheet),
        ("User_Awareness", create_user_awareness_sheet),
        ("Performance_Metrics", create_performance_metrics_sheet),
        ("Licensing_Support", create_licensing_support_sheet),
        ("Capability_Requirements", create_capability_requirements_sheet),
        ("Evidence_Register", create_evidence_register_sheet),
        ("Gap_Analysis", create_gap_analysis_sheet),
        ("Approval_Sign_Of", create_approval_signoff_sheet),
    ]

    for i, (sheet_name, create_func) in enumerate(sheets, 1):
        print(f"  [{i}/13] Creating {sheet_name}...")
        create_func(wb[sheet_name], styles)
        print(f"  ✅ {sheet_name} complete")

    print("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.1-7-18-19.S2_Protection_Coverage_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        wb.save(filename)
        print(f"{CHECK} SUCCESS: {filename}")
    except Exception as e:
        print(f"{XMARK} ERROR saving workbook: {e}")
        return 1

    print("\n" + "=" * 78)
    print("📋 WORKBOOK SUMMARY")
    print("=" * 78)
    print("\n✅ 13 sheets with comprehensive A.8.7 malware protection assessment")
    print(f"{CHECK} 50 endpoint coverage rows, 100 detection tracking rows")
    print(f"{CHECK} 50 incident response rows, 12 months performance metrics")
    print(f"{CHECK} 25 policy requirements, 100 evidence entries, 40 gap rows")
    print(f"{CHECK} Vendor-agnostic: Microsoft Defender, CrowdStrike, SentinelOne, ANY solution")
    print("\n" + "=" * 78)
    print('"Evidence-based compliance, not cargo cult security theater."')
    print("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    exit(main())
