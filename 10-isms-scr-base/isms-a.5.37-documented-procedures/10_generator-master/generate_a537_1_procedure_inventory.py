#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.37.1 - Procedure Inventory Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.37: Documented Operating Procedures
Assessment Domain 1 of 4: Procedure Inventory and Gap Analysis

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific procedure documentation requirements, categories,
and assessment criteria.

Key customization areas:
1. Procedure categories (adapt to your organizational structure)
2. Review cycle requirements (align with your governance framework)
3. Gap analysis criteria (based on your compliance requirements)
4. Role definitions (match your access control matrix)
5. Evidence requirements (per your audit standards)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-IMP-A.5.37.1 specification

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for cataloguing
and assessing documented operating procedures across information processing
facilities.

**Purpose:**
Enables systematic inventory of all operating procedures, tracking ownership,
accessibility, and currency status against ISO 27001:2022 Control A.5.37
requirements.

**Assessment Scope:**
- System operational procedures (backup, recovery, maintenance)
- Security operational procedures (incident response, access review)
- Facility operational procedures (HVAC, access controls, alarms)
- Change management procedures
- Recovery operations (DR, business continuity)
- User management procedures (onboarding, offboarding)

**Generated Workbook Structure:**
1. Instructions - Assessment guidance and methodology
2. Procedure_Inventory - Master catalogue of all documented procedures
3. Required_Procedures - ISO 27001 required procedure reference list
4. Accessibility_Matrix - Role-based procedure access mapping
5. Gap_Analysis - Missing/incomplete procedure tracking
6. Evidence_Register - Audit evidence tracking
7. Approval_SignOff - Stakeholder review and approval

================================================================================
"""

# =============================================================================
# IMPORTS - Standard Library
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# IMPORTS - Third Party
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.37.1"
WORKBOOK_NAME = "Procedure Inventory Assessment"
CONTROL_ID = "A.5.37"
CONTROL_NAME = "Documented Operating Procedures"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches specification
    sheets = [
        "Instructions",
        "Procedure_Inventory",
        "Required_Procedures",
        "Accessibility_Matrix",
        "Gap_Analysis",
        "Evidence_Register",
        "Approval_SignOff",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet with assessment guidance."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Procedure Inventory & Gap Analysis"),
        ("Control Reference", CONTROL_ID),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Annual"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Complete the Procedure_Inventory sheet with all documented procedures",
        "2. Review Required_Procedures list against your inventory",
        "3. Complete Accessibility_Matrix to verify role-based access",
        "4. Document gaps in Gap_Analysis sheet",
        "5. Link evidence in Evidence_Register",
        "6. Obtain approvals in Approval_SignOff sheet",
    ]

    row += 1
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    row += 1
    ws[f"A{row}"] = "PROCEDURE CATEGORIES"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    categories = [
        ("System Operations", "Core IT procedures: backup, restart, monitoring"),
        ("Security Operations", "Security-specific: incident response, access review"),
        ("Facility Operations", "Physical facility: HVAC, fire suppression, entry"),
        ("Change Management", "Change control: CAB process, emergency change"),
        ("Recovery Operations", "Business continuity: DR activation, failover"),
        ("User Management", "Identity lifecycle: onboarding, offboarding"),
    ]

    row += 1
    for cat, desc in categories:
        ws[f"A{row}"] = cat
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = desc
        row += 1

    row += 1
    ws[f"A{row}"] = "APPROVAL STATUS DEFINITIONS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    statuses = [
        ("Draft", "Document in development, not approved"),
        ("Pending Approval", "Submitted for review"),
        ("Approved", "Formally approved by management"),
        ("Expired", "Past review date, requires update"),
        ("Under Review", "Currently being revised"),
    ]

    row += 1
    for status, desc in statuses:
        ws[f"A{row}"] = status
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = desc
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    ws.freeze_panes = "A4"


# =============================================================================
# PROCEDURE INVENTORY SHEET
# =============================================================================
def create_procedure_inventory_sheet(ws, styles):
    """Create the Procedure_Inventory sheet - master catalogue."""
    ws.merge_cells("A1:P1")
    ws["A1"] = "PROCEDURE INVENTORY - Master Catalogue"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Column definitions per specification
    columns = [
        ("Procedure_ID", 18),
        ("Procedure_Name", 40),
        ("Category", 20),
        ("Process_Owner", 25),
        ("Department", 20),
        ("Document_Location", 35),
        ("Last_Review_Date", 16),
        ("Next_Review_Due", 16),
        ("Review_Cycle_Days", 16),
        ("Version", 12),
        ("Approval_Status", 18),
        ("Approver", 25),
        ("Approval_Date", 16),
        ("Related_Controls", 25),
        ("Criticality", 14),
        ("Notes", 35),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data validations
    dv_category = DataValidation(
        type="list",
        formula1='"System Operations,Security Operations,Facility Operations,Change Management,Recovery Operations,User Management,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_category)

    dv_status = DataValidation(
        type="list",
        formula1='"Draft,Pending Approval,Approved,Expired,Under Review"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_criticality = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_criticality)

    # Data entry rows (4-103)
    for r in range(4, 104):
        # Pre-fill Procedure_ID pattern
        ws.cell(row=r, column=1, value=f"SOP-XXX-{r-3:03d}").font = Font(color="808080")

        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Apply validations
        dv_category.add(ws.cell(row=r, column=3))  # Category
        dv_status.add(ws.cell(row=r, column=11))   # Approval_Status
        dv_criticality.add(ws.cell(row=r, column=15))  # Criticality

        # Next_Review_Due formula
        ws.cell(row=r, column=8).value = f"=IF(G{r}<>\"\",G{r}+I{r},\"\")"
        ws.cell(row=r, column=8).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Default review cycle
        ws.cell(row=r, column=9).value = 365

    ws.freeze_panes = "C4"


# =============================================================================
# REQUIRED PROCEDURES SHEET
# =============================================================================
def create_required_procedures_sheet(ws, styles):
    """Create the Required_Procedures sheet - ISO 27001 reference list."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "REQUIRED PROCEDURES - ISO 27001:2022 Reference List"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Reference_ID", 15),
        ("Required_Procedure", 50),
        ("ISO_Control", 20),
        ("Category", 22),
        ("Priority", 14),
        ("Current_Status", 18),
        ("Mapped_Procedure_ID", 20),
        ("Gap_Notes", 40),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate with common required procedures
    required_procedures = [
        ("REQ-001", "Information Security Incident Response Procedure", "A.5.24-28", "Security Operations", "High"),
        ("REQ-002", "Access Control Management Procedure", "A.5.15-18", "User Management", "High"),
        ("REQ-003", "Change Management Procedure", "A.8.32", "Change Management", "High"),
        ("REQ-004", "Backup and Recovery Procedure", "A.8.13", "System Operations", "Critical"),
        ("REQ-005", "Business Continuity Activation Procedure", "A.5.29-30", "Recovery Operations", "Critical"),
        ("REQ-006", "User Onboarding/Offboarding Procedure", "A.6.1-5", "User Management", "High"),
        ("REQ-007", "Vulnerability Management Procedure", "A.8.8", "Security Operations", "High"),
        ("REQ-008", "Logging and Monitoring Procedure", "A.8.15-16", "System Operations", "High"),
        ("REQ-009", "Physical Access Control Procedure", "A.7.1-4", "Facility Operations", "High"),
        ("REQ-010", "Media Handling and Disposal Procedure", "A.7.10, A.7.14", "Facility Operations", "Medium"),
        ("REQ-011", "Cryptographic Key Management Procedure", "A.8.24", "Security Operations", "High"),
        ("REQ-012", "Network Security Management Procedure", "A.8.20-22", "System Operations", "High"),
        ("REQ-013", "Software Development Security Procedure", "A.8.25-31", "System Operations", "High"),
        ("REQ-014", "Supplier Security Management Procedure", "A.5.19-23", "Security Operations", "Medium"),
        ("REQ-015", "Evidence Collection and Handling Procedure", "A.5.28", "Security Operations", "Medium"),
    ]

    # Data validations
    dv_status = DataValidation(
        type="list",
        formula1='"Exists,Partial,Missing"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_priority)

    # Populate rows
    for row_idx, (ref_id, proc, control, category, priority) in enumerate(required_procedures, start=4):
        ws.cell(row=row_idx, column=1, value=ref_id)
        ws.cell(row=row_idx, column=2, value=proc)
        ws.cell(row=row_idx, column=3, value=control)
        ws.cell(row=row_idx, column=4, value=category)
        ws.cell(row=row_idx, column=5, value=priority)

        for c in range(6, 9):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_status.add(ws.cell(row=row_idx, column=6))
        dv_priority.add(ws.cell(row=row_idx, column=5))

    # Additional blank rows
    for r in range(len(required_procedures) + 4, len(required_procedures) + 24):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_status.add(ws.cell(row=r, column=6))
        dv_priority.add(ws.cell(row=r, column=5))

    ws.freeze_panes = "C4"


# =============================================================================
# ACCESSIBILITY MATRIX SHEET
# =============================================================================
def create_accessibility_matrix_sheet(ws, styles):
    """Create the Accessibility_Matrix sheet - role-based access mapping."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "ACCESSIBILITY MATRIX - Role-Based Procedure Access"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Procedure_ID", 18),
        ("IT Operations", 16),
        ("Security Team", 16),
        ("Facilities", 16),
        ("Help Desk", 16),
        ("Management", 16),
        ("Access_Method", 25),
        ("Verified_Date", 16),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Boolean dropdown for role columns
    dv_bool = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_bool)

    # Data entry rows
    for r in range(4, 104):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Apply boolean validation to role columns (2-6)
        for c in range(2, 7):
            dv_bool.add(ws.cell(row=r, column=c))

    ws.freeze_panes = "B4"


# =============================================================================
# GAP ANALYSIS SHEET
# =============================================================================
def create_gap_analysis_sheet(ws, styles):
    """Create the Gap_Analysis sheet - track procedure gaps and remediation."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "GAP ANALYSIS - Procedure Gaps and Remediation"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Gap_ID", 12),
        ("Gap_Type", 18),
        ("Procedure_Reference", 35),
        ("Severity", 14),
        ("Identified_Date", 16),
        ("Remediation_Owner", 25),
        ("Target_Date", 16),
        ("Status", 16),
        ("Completion_Date", 16),
        ("Evidence", 35),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Validations
    dv_gap_type = DataValidation(
        type="list",
        formula1='"Missing,Incomplete,Outdated,Unapproved"',
        allow_blank=False
    )
    ws.add_data_validation(dv_gap_type)

    dv_severity = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_severity)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Closed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    # Data entry rows
    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"GAP-{r-3:03d}").font = Font(color="808080")

        for c in range(2, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_gap_type.add(ws.cell(row=r, column=2))
        dv_severity.add(ws.cell(row=r, column=4))
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "C4"


# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================
def create_evidence_register_sheet(ws, styles):
    """Create the Evidence_Register sheet for audit traceability."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 45),
        ("Related_Procedure", 25),
        ("Collection_Date", 16),
        ("Location", 40),
        ("Collected_By", 25),
        ("Valid_Until", 16),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Evidence type validation
    dv_type = DataValidation(
        type="list",
        formula1='"Document,Screenshot,Export,Attestation,Configuration,Audit Log"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    # Data entry rows
    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")

        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "C4"


# =============================================================================
# APPROVAL SIGN-OFF SHEET
# =============================================================================
def create_approval_signoff_sheet(ws, styles):
    """Create the Approval_SignOff sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Total Procedures Inventoried", "=COUNTA(Procedure_Inventory!A4:A103)-COUNTBLANK(Procedure_Inventory!B4:B103)"),
        ("Procedures Approved", "=COUNTIF(Procedure_Inventory!K4:K103,\"Approved\")"),
        ("Open Gaps", "=COUNTIF(Gap_Analysis!H4:H53,\"Open\")"),
        ("Assessment Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "" or not value.startswith("="):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    # Assessment Completed By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    completion_fields = ["Name", "Role/Title", "Department", "Email", "Date"]
    row += 1
    for field in completion_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Reviewed By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (Information Security Manager)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    review_fields = ["Name", "Date", "Signature"]
    row += 1
    for field in review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Approved By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    approval_fields = ["Name", "Date", "Approval Decision", "Signature"]
    row += 1
    for field in approval_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Decision dropdown
    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add(ws[f"B{row-2}"])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A3"


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main() -> int:
    """
    Main execution function - orchestrates workbook creation.

    Returns:
        int: 0 on success, 1 on failure
    """
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/7] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/7] Creating Procedure_Inventory sheet...")
        create_procedure_inventory_sheet(wb["Procedure_Inventory"], styles)

        logger.info("[3/7] Creating Required_Procedures sheet...")
        create_required_procedures_sheet(wb["Required_Procedures"], styles)

        logger.info("[4/7] Creating Accessibility_Matrix sheet...")
        create_accessibility_matrix_sheet(wb["Accessibility_Matrix"], styles)

        logger.info("[5/7] Creating Gap_Analysis sheet...")
        create_gap_analysis_sheet(wb["Gap_Analysis"], styles)

        logger.info("[6/7] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[7/7] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("Next steps:")
        logger.info("  1) Complete document information in Instructions")
        logger.info("  2) Inventory all procedures in Procedure_Inventory")
        logger.info("  3) Map required procedures against inventory")
        logger.info("  4) Complete Accessibility_Matrix")
        logger.info("  5) Document gaps in Gap_Analysis")
        logger.info("  6) Link evidence in Evidence_Register")
        logger.info("  7) Obtain approvals")
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        logger.error("Install with: pip install openpyxl")
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# END OF SCRIPT
# =============================================================================
# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.37.1 specification
# =============================================================================
