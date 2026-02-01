#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.37.3 - Procedure Review and Update Tracking Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.37: Documented Operating Procedures
Assessment Domain 3 of 4: Review Cycle and Change Management

Reference Pattern: Based on ISMS-IMP-A.5.37.3 specification
================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.37.3"
WORKBOOK_NAME = "Procedure Review and Update Tracking"
CONTROL_ID = "A.5.37"
CONTROL_NAME = "Documented Operating Procedures"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "overdue": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "due_soon": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "current": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
    }
    return styles


def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Review_Schedule",
        "Change_Requests",
        "Version_History",
        "Communication_Log",
        "Overdue_Escalation",
        "Evidence_Register",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Review Cycle & Change Management"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "MANDATORY REVIEW TRIGGERS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    triggers = [
        ("Scheduled Review", "Per review cycle (annual default)"),
        ("System Change", "Within 30 days of change"),
        ("Incident Related", "Within 14 days of incident"),
        ("Regulatory Change", "Within 60 days of effective date"),
        ("Audit Finding", "Per CAR timeline"),
        ("Personnel Change", "Within 30 days"),
    ]

    row += 1
    for trigger, timeframe in triggers:
        ws[f"A{row}"] = trigger
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = timeframe
        row += 1

    row += 1
    ws[f"A{row}"] = "REVIEW CYCLE BY CRITICALITY"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    cycles = [
        ("Critical", "6 months"),
        ("High", "12 months"),
        ("Medium", "12 months"),
        ("Low", "24 months"),
    ]

    row += 1
    for crit, cycle in cycles:
        ws[f"A{row}"] = crit
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = cycle
        row += 1

    row += 1
    ws[f"A{row}"] = "ESCALATION MATRIX"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    escalation = [
        ("L1 (1-14 days overdue)", "Procedure Owner + Line Manager"),
        ("L2 (15-30 days overdue)", "Department Head + ISM"),
        ("L3 (>30 days overdue)", "CISO + Executive Management"),
    ]

    row += 1
    for level, recipient in escalation:
        ws[f"A{row}"] = level
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = recipient
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 45


def create_review_schedule_sheet(ws, styles):
    """Create the Review_Schedule sheet."""
    ws.merge_cells("A1:N1")
    ws["A1"] = "REVIEW SCHEDULE - Procedure Review Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Procedure_ID", 18),
        ("Procedure_Name", 40),
        ("Criticality", 14),
        ("Review_Cycle_Days", 16),
        ("Last_Review_Date", 16),
        ("Next_Review_Due", 16),
        ("Days_Until_Due", 14),
        ("Review_Status", 16),
        ("Assigned_Reviewer", 22),
        ("Review_Started", 16),
        ("Review_Completed", 16),
        ("Review_Outcome", 20),
        ("New_Version", 14),
        ("Notes", 35),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_criticality = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_criticality)

    dv_outcome = DataValidation(
        type="list",
        formula1='"Current - No Changes,Minor Updates,Major Updates,Obsolete,Superseded"',
        allow_blank=False
    )
    ws.add_data_validation(dv_outcome)

    for r in range(4, 104):
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_criticality.add(ws.cell(row=r, column=3))
        dv_outcome.add(ws.cell(row=r, column=12))

        # Next_Review_Due = Last_Review_Date + Review_Cycle_Days
        ws.cell(row=r, column=6).value = f"=IF(E{r}<>\"\",E{r}+D{r},\"\")"
        ws.cell(row=r, column=6).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Days_Until_Due
        ws.cell(row=r, column=7).value = f"=IF(F{r}<>\"\",F{r}-TODAY(),\"\")"
        ws.cell(row=r, column=7).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Review_Status
        ws.cell(row=r, column=8).value = f'=IF(G{r}="","",IF(G{r}<0,"OVERDUE",IF(G{r}<=30,"DUE SOON","CURRENT")))'
        ws.cell(row=r, column=8).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.freeze_panes = "C4"


def create_change_requests_sheet(ws, styles):
    """Create the Change_Requests sheet."""
    ws.merge_cells("A1:N1")
    ws["A1"] = "CHANGE REQUESTS - Procedure Change Management"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("CR_ID", 18),
        ("Procedure_ID", 18),
        ("Request_Date", 14),
        ("Requestor", 22),
        ("Change_Category", 18),
        ("Trigger", 18),
        ("Description", 40),
        ("Justification", 35),
        ("Impact_Assessment", 35),
        ("Status", 16),
        ("Approver", 22),
        ("Approval_Date", 14),
        ("Implementation_Date", 16),
        ("Verification", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_category = DataValidation(
        type="list",
        formula1='"Administrative,Minor,Major,Emergency"',
        allow_blank=False
    )
    ws.add_data_validation(dv_category)

    dv_trigger = DataValidation(
        type="list",
        formula1='"Scheduled Review,System Change,Incident Related,Regulatory Change,Audit Finding,Personnel Change"',
        allow_blank=False
    )
    ws.add_data_validation(dv_trigger)

    dv_status = DataValidation(
        type="list",
        formula1='"Submitted,Under Review,Approved,Rejected,Implemented"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        # Auto-generate CR_ID
        ws.cell(row=r, column=1, value=f"CR-{datetime.now().strftime('%Y%m')}-{r-3:03d}").font = Font(color="808080")

        for c in range(2, 15):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_category.add(ws.cell(row=r, column=5))
        dv_trigger.add(ws.cell(row=r, column=6))
        dv_status.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "C4"


def create_version_history_sheet(ws, styles):
    """Create the Version_History sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "VERSION HISTORY - Procedure Version Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Procedure_ID", 18),
        ("Version", 12),
        ("Effective_Date", 16),
        ("Supersedes", 14),
        ("Change_Summary", 50),
        ("CR_Reference", 20),
        ("Approved_By", 25),
        ("Status", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_status = DataValidation(
        type="list",
        formula1='"Active,Superseded,Archived"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "C4"


def create_communication_log_sheet(ws, styles):
    """Create the Communication_Log sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "COMMUNICATION LOG - Procedure Update Communications"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Comm_ID", 14),
        ("Procedure_ID", 18),
        ("Version", 12),
        ("Communication_Date", 18),
        ("Communication_Method", 20),
        ("Audience", 30),
        ("Acknowledgement_Required", 22),
        ("Acknowledgement_Rate", 18),
        ("Training_Required", 16),
        ("Training_Completion", 18),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_method = DataValidation(
        type="list",
        formula1='"Email,Intranet,Meeting,Training"',
        allow_blank=False
    )
    ws.add_data_validation(dv_method)

    dv_bool = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_bool)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"COMM-{r-3:03d}").font = Font(color="808080")

        for c in range(2, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_method.add(ws.cell(row=r, column=5))
        dv_bool.add(ws.cell(row=r, column=7))
        dv_bool.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "C4"


def create_overdue_escalation_sheet(ws, styles):
    """Create the Overdue_Escalation sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "OVERDUE ESCALATION - Review Escalation Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Procedure_ID", 18),
        ("Days_Overdue", 14),
        ("Escalation_Level", 16),
        ("Escalated_To", 30),
        ("Escalation_Date", 16),
        ("Response_Required_By", 18),
        ("Status", 16),
        ("Resolution", 40),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_level = DataValidation(
        type="list",
        formula1='"L1,L2,L3"',
        allow_blank=False
    )
    ws.add_data_validation(dv_level)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,Acknowledged,Resolved"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_level.add(ws.cell(row=r, column=3))
        dv_status.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "B4"


def create_evidence_register_sheet(ws, styles):
    """Create the Evidence_Register sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 25),
        ("Related_CR", 18),
        ("Related_Procedure", 22),
        ("Description", 45),
        ("Collection_Date", 16),
        ("Location", 40),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Review Record,Approval,Communication,Training"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 8):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "C4"


def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/7] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/7] Creating Review_Schedule sheet...")
        create_review_schedule_sheet(wb["Review_Schedule"], styles)

        logger.info("[3/7] Creating Change_Requests sheet...")
        create_change_requests_sheet(wb["Change_Requests"], styles)

        logger.info("[4/7] Creating Version_History sheet...")
        create_version_history_sheet(wb["Version_History"], styles)

        logger.info("[5/7] Creating Communication_Log sheet...")
        create_communication_log_sheet(wb["Communication_Log"], styles)

        logger.info("[6/7] Creating Overdue_Escalation sheet...")
        create_overdue_escalation_sheet(wb["Overdue_Escalation"], styles)

        logger.info("[7/7] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.37.3 specification
# =============================================================================
