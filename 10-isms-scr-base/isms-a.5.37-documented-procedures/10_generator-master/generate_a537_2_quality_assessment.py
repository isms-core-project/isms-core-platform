#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.37.2 - Procedure Quality Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.37: Documented Operating Procedures
Assessment Domain 2 of 4: Procedure Quality Evaluation

Reference Pattern: Based on ISMS-IMP-A.5.37.2 specification
================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.37.2"
WORKBOOK_NAME = "Procedure Quality Assessment"
CONTROL_ID = "A.5.37"
CONTROL_NAME = "Documented Operating Procedures"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "score_excellent": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "score_good": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "score_poor": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Quality_Assessment",
        "Quality_Checklist",
        "Improvement_Actions",
        "Trend_Analysis",
        "Evidence_Register",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Procedure Quality Evaluation"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "QUALITY DIMENSIONS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    dimensions = [
        ("Clarity (20%)", "Can be understood by target audience"),
        ("Completeness (20%)", "Covers all necessary steps and scenarios"),
        ("Accuracy (20%)", "Reflects current systems and processes"),
        ("Usability (20%)", "Practical to follow in operations"),
        ("Maintainability (20%)", "Easy to update when changes occur"),
    ]

    row += 1
    for dim, desc in dimensions:
        ws[f"A{row}"] = dim
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = desc
        row += 1

    row += 1
    ws[f"A{row}"] = "QUALITY SCORING (0-5)"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    scores = [
        ("5 - Excellent", "Fully meets all criteria, best practice"),
        ("4 - Good", "Meets criteria with minor improvements needed"),
        ("3 - Adequate", "Meets minimum requirements"),
        ("2 - Needs Improvement", "Significant gaps exist"),
        ("1 - Poor", "Major deficiencies, high risk"),
        ("0 - Missing", "Element not present"),
    ]

    row += 1
    for score, desc in scores:
        ws[f"A{row}"] = score
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = desc
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55


def create_quality_assessment_sheet(ws, styles):
    """Create the Quality_Assessment sheet - main scoring."""
    ws.merge_cells("A1:N1")
    ws["A1"] = "QUALITY ASSESSMENT - Procedure Scoring"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Procedure_ID", 18),
        ("Procedure_Name", 40),
        ("Assessment_Date", 16),
        ("Assessor", 22),
        ("Clarity_Score", 14),
        ("Completeness_Score", 16),
        ("Accuracy_Score", 14),
        ("Usability_Score", 14),
        ("Maintainability_Score", 18),
        ("Overall_Score", 14),
        ("Quality_Rating", 16),
        ("Priority_Improvements", 30),
        ("Findings", 40),
        ("Next_Review", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Score validation (0-5)
    dv_score = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="5",
        allow_blank=True
    )
    dv_score.error = "Score must be between 0 and 5"
    ws.add_data_validation(dv_score)

    for r in range(4, 104):
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Apply score validation (columns E-I, indexes 5-9)
        for c in range(5, 10):
            dv_score.add(ws.cell(row=r, column=c))

        # Overall Score formula (weighted average)
        ws.cell(row=r, column=10).value = f"=IF(E{r}<>\"\",ROUND((E{r}*0.2)+(F{r}*0.2)+(G{r}*0.2)+(H{r}*0.2)+(I{r}*0.2),2),\"\")"
        ws.cell(row=r, column=10).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Quality Rating formula
        ws.cell(row=r, column=11).value = f'=IF(J{r}="","",IF(J{r}>=4.5,"Excellent",IF(J{r}>=3.5,"Good",IF(J{r}>=2.5,"Adequate",IF(J{r}>=1.5,"Needs Improvement","Poor")))))'
        ws.cell(row=r, column=11).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Next Review formula (based on rating)
        ws.cell(row=r, column=14).value = f'=IF(K{r}="Excellent",C{r}+365,IF(K{r}="Good",C{r}+180,IF(K{r}="Adequate",C{r}+90,IF(K{r}<>"",C{r}+30,""))))'
        ws.cell(row=r, column=14).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.freeze_panes = "C4"


def create_quality_checklist_sheet(ws, styles):
    """Create the Quality_Checklist sheet - detailed checklist items."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "QUALITY CHECKLIST - Detailed Assessment Items"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Procedure_ID", 18),
        ("Check_Category", 20),
        ("Check_Item", 55),
        ("Status", 14),
        ("Finding", 40),
        ("Recommendation", 40),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate checklist items
    checklist_items = [
        ("Document Structure", "Document ID and title present"),
        ("Document Structure", "Version control information"),
        ("Document Structure", "Owner and approver identified"),
        ("Document Structure", "Last review date documented"),
        ("Document Structure", "Scope clearly defined"),
        ("Document Structure", "Prerequisites listed"),
        ("Document Structure", "Related documents referenced"),
        ("Content Quality", "Purpose statement clear"),
        ("Content Quality", "Step-by-step instructions"),
        ("Content Quality", "Decision trees/flowcharts where needed"),
        ("Content Quality", "Screenshots/diagrams included"),
        ("Content Quality", "Error handling documented"),
        ("Content Quality", "Escalation procedures defined"),
        ("Content Quality", "Contact information current"),
        ("Content Quality", "Glossary of terms (if needed)"),
        ("Operational Elements", "Pre-execution checks defined"),
        ("Operational Elements", "Execution steps numbered"),
        ("Operational Elements", "Verification steps included"),
        ("Operational Elements", "Expected outcomes stated"),
        ("Operational Elements", "Rollback procedures documented"),
        ("Operational Elements", "Recovery steps included"),
        ("Operational Elements", "Time estimates provided"),
        ("Operational Elements", "Resource requirements listed"),
    ]

    dv_status = DataValidation(
        type="list",
        formula1='"Pass,Partial,Fail,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    row = 4
    for category, item in checklist_items:
        ws.cell(row=row, column=1).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=1).border = styles["border"]

        ws.cell(row=row, column=2, value=category)
        ws.cell(row=row, column=2).border = styles["border"]

        ws.cell(row=row, column=3, value=item)
        ws.cell(row=row, column=3).border = styles["border"]

        for c in range(4, 7):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_status.add(ws.cell(row=row, column=4))
        row += 1

    ws.freeze_panes = "B4"


def create_improvement_actions_sheet(ws, styles):
    """Create the Improvement_Actions sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "IMPROVEMENT ACTIONS - Quality Remediation Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Action_ID", 12),
        ("Procedure_ID", 18),
        ("Dimension", 18),
        ("Issue_Description", 40),
        ("Action_Required", 40),
        ("Owner", 22),
        ("Priority", 14),
        ("Target_Date", 16),
        ("Status", 16),
        ("Completion_Date", 16),
        ("Verification", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_dimension = DataValidation(
        type="list",
        formula1='"Clarity,Completeness,Accuracy,Usability,Maintainability"',
        allow_blank=False
    )
    ws.add_data_validation(dv_dimension)

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_priority)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Completed,Cancelled"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"ACT-{r-3:03d}").font = Font(color="808080")

        for c in range(2, 12):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_dimension.add(ws.cell(row=r, column=3))
        dv_priority.add(ws.cell(row=r, column=7))
        dv_status.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "C4"


def create_trend_analysis_sheet(ws, styles):
    """Create the Trend_Analysis sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "TREND ANALYSIS - Quality Improvement Over Time"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Period", 16),
        ("Procedures_Assessed", 18),
        ("Avg_Clarity", 14),
        ("Avg_Completeness", 16),
        ("Avg_Accuracy", 14),
        ("Avg_Usability", 14),
        ("Avg_Maintainability", 18),
        ("Overall_Avg", 14),
        ("Excellent_Count", 16),
        ("Poor_Count", 14),
        ("Improvement_%", 14),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(4, 24):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    ws.freeze_panes = "B4"


def create_evidence_register_sheet(ws, styles):
    """Create the Evidence_Register sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Procedure_ID", 18),
        ("Description", 45),
        ("Collection_Date", 16),
        ("Location", 40),
        ("Collected_By", 25),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Assessment,Screenshot,Document"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 8):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "C4"


def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/6] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/6] Creating Quality_Assessment sheet...")
        create_quality_assessment_sheet(wb["Quality_Assessment"], styles)

        logger.info("[3/6] Creating Quality_Checklist sheet...")
        create_quality_checklist_sheet(wb["Quality_Checklist"], styles)

        logger.info("[4/6] Creating Improvement_Actions sheet...")
        create_improvement_actions_sheet(wb["Improvement_Actions"], styles)

        logger.info("[5/6] Creating Trend_Analysis sheet...")
        create_trend_analysis_sheet(wb["Trend_Analysis"], styles)

        logger.info("[6/6] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.37.2 specification
# =============================================================================
