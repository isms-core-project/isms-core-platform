#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
A.5.14.2 Channel Security Assessment Generator
================================================================================

Generates Excel workbook for assessing security of information transfer channels
including email, cloud services, file sharing, and physical methods per ISO 27001:2022.

Sheets:
    1. Instructions - Completion guidance
    2. Email_Assessment - Corporate email security evaluation
    3. Cloud_Services - Cloud storage and sharing assessment
    4. File_Transfer - FTP, SFTP, MFT systems assessment
    5. Physical_Channels - Physical transfer security evaluation
    6. Risk_Assessment - Channel-specific risk analysis
    7. Evidence_Register - Supporting documentation
    8. Approval_SignOff - Authorization workflow

Usage:
    python3 generate_a514_2_channel_security_assessment.py

Output:
    ISMS-IMP-A.5.14.2_Channel_Security_Assessment_YYYYMMDD.xlsx

================================================================================
"""

import logging
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.14.2"
WORKBOOK_NAME = "Channel Security Assessment"
CONTROL_ID = "A.5.14"
CONTROL_NAME = "Information Transfer"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP_LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================
def set_column_widths(ws, widths: dict):
    """Set column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_header_row(ws, row: int, headers: list, fill=SUBHEADER_FILL, font=SUBHEADER_FONT):
    """Create a styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def add_data_validation(ws, cell_range: str, formula: str):
    """Add dropdown data validation."""
    dv = DataValidation(type="list", formula1=formula, showDropDown=False, allowBlank=True)
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


# =============================================================================
# SHEET CREATORS
# =============================================================================
def create_instructions_sheet(wb: Workbook):
    """Create the Instructions sheet."""
    ws = wb.active
    ws.title = "Instructions"

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Metadata
    metadata = [
        ("Document ID:", DOCUMENT_ID),
        ("Control Reference:", CONTROL_REF),
        ("Generated Date:", GENERATED_DATE),
        ("Version:", "1.0"),
        ("Classification:", "INTERNAL"),
    ]

    row = 3
    for label, value in metadata:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        row += 1

    # Purpose
    row += 1
    ws.cell(row=row, column=1, value="PURPOSE").font = BOLD_FONT
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    purpose_cell = ws.cell(row=row, column=1)
    purpose_cell.value = (
        "This workbook provides a structured assessment of information transfer channel security. "
        "It evaluates encryption, authentication, logging, and access controls for each transfer method "
        "used by the organization. Complete this assessment to identify security gaps and prioritize "
        "remediation efforts for transfer channels."
    )
    purpose_cell.font = NORMAL_FONT
    purpose_cell.alignment = TOP_LEFT_ALIGN
    ws.row_dimensions[row].height = 60

    # Sheet descriptions
    row += 2
    ws.cell(row=row, column=1, value="SHEET DESCRIPTIONS").font = BOLD_FONT
    row += 1

    sheets = [
        ("Email_Assessment", "Security evaluation of email systems and configurations"),
        ("Cloud_Services", "Assessment of cloud storage and sharing platforms"),
        ("File_Transfer", "Evaluation of FTP, SFTP, MFT, and API transfer systems"),
        ("Physical_Channels", "Assessment of physical media and courier transfer security"),
        ("Risk_Assessment", "Channel-specific risk analysis and scoring"),
        ("Evidence_Register", "Supporting documentation and evidence tracking"),
        ("Approval_SignOff", "Assessment approval workflow"),
    ]

    for sheet_name, description in sheets:
        ws.cell(row=row, column=1, value=sheet_name).font = BOLD_FONT
        ws.cell(row=row, column=2, value=description).font = NORMAL_FONT
        row += 1

    # Assessment rating legend
    row += 1
    ws.cell(row=row, column=1, value="ASSESSMENT RATINGS").font = BOLD_FONT
    row += 1

    ratings = [
        ("Compliant", PASS_FILL, "Control fully implemented and effective"),
        ("Partial", PARTIAL_FILL, "Control partially implemented, gaps exist"),
        ("Non-Compliant", FAIL_FILL, "Control not implemented or ineffective"),
        ("N/A", None, "Control not applicable to this channel"),
    ]

    for rating, fill, desc in ratings:
        cell = ws.cell(row=row, column=1, value=rating)
        if fill:
            cell.fill = fill
        cell.font = BOLD_FONT
        ws.cell(row=row, column=2, value=desc).font = NORMAL_FONT
        row += 1

    set_column_widths(ws, {"A": 25, "B": 60, "C": 20, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20})
    ws.freeze_panes = "A3"

    logger.info("Created Instructions sheet")


def create_email_assessment_sheet(wb: Workbook):
    """Create the Email Assessment sheet."""
    ws = wb.create_sheet("Email_Assessment")

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "Email Security Assessment"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Control Area", "Security Requirement", "Expected Configuration", "Actual Configuration", "Status", "Evidence", "Gap Description", "Remediation", "Priority"]
    create_header_row(ws, 3, headers)

    # Assessment items
    items = [
        ("Transport Encryption", "TLS enforcement for all email", "TLS 1.2+ required, opportunistic TLS for external", "", "", "", "", "", ""),
        ("Transport Encryption", "DANE/MTA-STS configured", "MTA-STS policy published, DANE if supported", "", "", "", "", "", ""),
        ("Message Encryption", "S/MIME or PGP available", "S/MIME certificates issued to all users", "", "", "", "", "", ""),
        ("Message Encryption", "Automatic encryption for sensitive", "DLP rules trigger encryption for CONFIDENTIAL", "", "", "", "", "", ""),
        ("Authentication", "SPF record configured", "SPF -all or ~all published", "", "", "", "", "", ""),
        ("Authentication", "DKIM signing enabled", "DKIM 2048-bit keys, selector rotation", "", "", "", "", "", ""),
        ("Authentication", "DMARC policy enforced", "DMARC p=reject for primary domain", "", "", "", "", "", ""),
        ("Anti-Malware", "Attachment scanning", "All attachments scanned, macros blocked", "", "", "", "", "", ""),
        ("Anti-Malware", "URL rewriting/sandboxing", "Malicious URLs blocked, sandboxing enabled", "", "", "", "", "", ""),
        ("Data Loss Prevention", "DLP policies active", "Classification-based rules enforced", "", "", "", "", "", ""),
        ("Data Loss Prevention", "External forwarding controls", "Auto-forward to external blocked", "", "", "", "", "", ""),
        ("Logging & Monitoring", "Message tracking enabled", "90-day retention for message traces", "", "", "", "", "", ""),
        ("Logging & Monitoring", "Audit logging active", "Admin and mailbox audit logs enabled", "", "", "", "", "", ""),
        ("Access Control", "MFA enforced", "All accounts require MFA", "", "", "", "", "", ""),
        ("Access Control", "Conditional Access policies", "Location/device-based restrictions", "", "", "", "", "", ""),
    ]

    row = 4
    for item in items:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [4, 5, 6, 7, 8, 9]:  # Input fields
                cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"E4:E{row+5}", '"Compliant,Partial,Non-Compliant,N/A"')
    add_data_validation(ws, f"I4:I{row+5}", '"Critical,High,Medium,Low"')

    set_column_widths(ws, {"A": 20, "B": 28, "C": 35, "D": 30, "E": 15, "F": 20, "G": 30, "H": 30, "I": 12})
    ws.freeze_panes = "A4"

    logger.info("Created Email_Assessment sheet")


def create_cloud_services_sheet(wb: Workbook):
    """Create the Cloud Services assessment sheet."""
    ws = wb.create_sheet("Cloud_Services")

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "Cloud Services Security Assessment"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Service", "Control Area", "Requirement", "Configuration", "Status", "Evidence", "Gap", "Remediation", "Priority"]
    create_header_row(ws, 3, headers)

    # Assessment items
    items = [
        ("SharePoint/OneDrive", "Encryption at Rest", "AES-256 encryption", "", "", "", "", "", ""),
        ("SharePoint/OneDrive", "Encryption in Transit", "TLS 1.2+ enforced", "", "", "", "", "", ""),
        ("SharePoint/OneDrive", "Access Control", "Sensitivity labels applied", "", "", "", "", "", ""),
        ("SharePoint/OneDrive", "External Sharing", "Restricted to approved domains", "", "", "", "", "", ""),
        ("SharePoint/OneDrive", "Audit Logging", "Unified audit log enabled", "", "", "", "", "", ""),
        ("Teams", "Guest Access", "Guest access policy defined", "", "", "", "", "", ""),
        ("Teams", "File Sharing", "External file sharing restricted", "", "", "", "", "", ""),
        ("Teams", "Data Retention", "Retention policies applied", "", "", "", "", "", ""),
        ("Box/Dropbox", "SSO Integration", "SAML SSO configured", "", "", "", "", "", ""),
        ("Box/Dropbox", "DLP Integration", "CASB or native DLP active", "", "", "", "", "", ""),
        ("Box/Dropbox", "Device Trust", "Only managed devices allowed", "", "", "", "", "", ""),
        ("AWS S3", "Bucket Encryption", "Server-side encryption (SSE-S3/KMS)", "", "", "", "", "", ""),
        ("AWS S3", "Access Logging", "S3 access logging enabled", "", "", "", "", "", ""),
        ("AWS S3", "Public Access Block", "Public access blocked by default", "", "", "", "", "", ""),
        ("Azure Blob", "Encryption", "Storage service encryption enabled", "", "", "", "", "", ""),
        ("Azure Blob", "Network Security", "Private endpoints or service endpoints", "", "", "", "", "", ""),
        ("Google Drive", "Sharing Settings", "External sharing restrictions", "", "", "", "", "", ""),
        ("Google Drive", "DLP", "Google Workspace DLP rules active", "", "", "", "", "", ""),
    ]

    row = 4
    for item in items:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [4, 5, 6, 7, 8, 9]:  # Input fields
                cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"E4:E{row+5}", '"Compliant,Partial,Non-Compliant,N/A"')
    add_data_validation(ws, f"I4:I{row+5}", '"Critical,High,Medium,Low"')

    set_column_widths(ws, {"A": 20, "B": 20, "C": 30, "D": 30, "E": 15, "F": 18, "G": 25, "H": 25, "I": 12})
    ws.freeze_panes = "A4"

    logger.info("Created Cloud_Services sheet")


def create_file_transfer_sheet(wb: Workbook):
    """Create the File Transfer systems assessment sheet."""
    ws = wb.create_sheet("File_Transfer")

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "File Transfer Systems Security Assessment"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["System Type", "Control", "Requirement", "Current State", "Status", "Evidence", "Gap", "Remediation", "Priority"]
    create_header_row(ws, 3, headers)

    # Assessment items
    items = [
        ("SFTP Server", "Protocol Version", "SSH 2.0 only, deprecated algorithms disabled", "", "", "", "", "", ""),
        ("SFTP Server", "Authentication", "Key-based auth required, password disabled", "", "", "", "", "", ""),
        ("SFTP Server", "Access Control", "Chroot jails per user/group", "", "", "", "", "", ""),
        ("SFTP Server", "Logging", "All transfers logged with user/file/timestamp", "", "", "", "", "", ""),
        ("SFTP Server", "Encryption", "AES-256-CTR or AES-256-GCM ciphers", "", "", "", "", "", ""),
        ("MFT Platform", "Encryption at Rest", "Files encrypted with AES-256", "", "", "", "", "", ""),
        ("MFT Platform", "Encryption in Transit", "TLS 1.2+ or SSH", "", "", "", "", "", ""),
        ("MFT Platform", "Workflow Automation", "Automated routing based on classification", "", "", "", "", "", ""),
        ("MFT Platform", "Non-Repudiation", "Digital signatures for transfers", "", "", "", "", "", ""),
        ("MFT Platform", "Audit Trail", "Complete audit trail with retention", "", "", "", "", "", ""),
        ("API Gateway", "Authentication", "OAuth 2.0 / mTLS configured", "", "", "", "", "", ""),
        ("API Gateway", "Rate Limiting", "Rate limits per client enforced", "", "", "", "", "", ""),
        ("API Gateway", "Input Validation", "Schema validation enabled", "", "", "", "", "", ""),
        ("API Gateway", "Logging", "All requests logged to SIEM", "", "", "", "", "", ""),
        ("Legacy FTP", "Encryption", "FTPS (explicit) or FTP over VPN only", "", "", "", "", "", ""),
        ("Legacy FTP", "Network Isolation", "FTP servers in isolated VLAN", "", "", "", "", "", ""),
        ("Legacy FTP", "Deprecation Plan", "Migration timeline to secure alternatives", "", "", "", "", "", ""),
    ]

    row = 4
    for item in items:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [4, 5, 6, 7, 8, 9]:  # Input fields
                cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"E4:E{row+5}", '"Compliant,Partial,Non-Compliant,N/A"')
    add_data_validation(ws, f"I4:I{row+5}", '"Critical,High,Medium,Low"')

    set_column_widths(ws, {"A": 18, "B": 20, "C": 35, "D": 30, "E": 15, "F": 18, "G": 25, "H": 25, "I": 12})
    ws.freeze_panes = "A4"

    logger.info("Created File_Transfer sheet")


def create_physical_channels_sheet(wb: Workbook):
    """Create the Physical Channels assessment sheet."""
    ws = wb.create_sheet("Physical_Channels")

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "Physical Transfer Channels Security Assessment"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Channel", "Control Area", "Requirement", "Current Practice", "Status", "Evidence", "Gap", "Remediation", "Priority"]
    create_header_row(ws, 3, headers)

    # Assessment items
    items = [
        ("USB/Removable Media", "Device Encryption", "Hardware encrypted USB only (FIPS 140-2)", "", "", "", "", "", ""),
        ("USB/Removable Media", "Device Management", "Approved device list enforced via DLP", "", "", "", "", "", ""),
        ("USB/Removable Media", "Malware Scanning", "Auto-scan on insertion", "", "", "", "", "", ""),
        ("USB/Removable Media", "Logging", "All USB connections logged", "", "", "", "", "", ""),
        ("Courier Services", "Approved Vendors", "Vetted courier list maintained", "", "", "", "", "", ""),
        ("Courier Services", "Tracking", "Real-time shipment tracking required", "", "", "", "", "", ""),
        ("Courier Services", "Packaging Standards", "Tamper-evident packaging for CONFIDENTIAL+", "", "", "", "", "", ""),
        ("Courier Services", "Chain of Custody", "Documented handoff procedures", "", "", "", "", "", ""),
        ("Internal Mail", "Envelope Security", "Sealed envelopes for INTERNAL+", "", "", "", "", "", ""),
        ("Internal Mail", "Delivery Confirmation", "Signature required for CONFIDENTIAL+", "", "", "", "", "", ""),
        ("Backup Media", "Encryption", "All backup tapes encrypted", "", "", "", "", "", ""),
        ("Backup Media", "Inventory Tracking", "Barcode tracking system active", "", "", "", "", "", ""),
        ("Backup Media", "Secure Storage", "Fireproof, access-controlled vault", "", "", "", "", "", ""),
        ("Backup Media", "Transport", "Secure courier for offsite rotation", "", "", "", "", "", ""),
        ("Print/Fax", "Secure Print", "PIN release for CONFIDENTIAL+", "", "", "", "", "", ""),
        ("Print/Fax", "Fax Transmission", "Encrypted fax or fax-to-email only", "", "", "", "", "", ""),
        ("Print/Fax", "Shredding", "Cross-cut shredding for INTERNAL+", "", "", "", "", "", ""),
    ]

    row = 4
    for item in items:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [4, 5, 6, 7, 8, 9]:  # Input fields
                cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"E4:E{row+5}", '"Compliant,Partial,Non-Compliant,N/A"')
    add_data_validation(ws, f"I4:I{row+5}", '"Critical,High,Medium,Low"')

    set_column_widths(ws, {"A": 20, "B": 20, "C": 35, "D": 30, "E": 15, "F": 18, "G": 25, "H": 25, "I": 12})
    ws.freeze_panes = "A4"

    logger.info("Created Physical_Channels sheet")


def create_risk_assessment_sheet(wb: Workbook):
    """Create the Risk Assessment sheet."""
    ws = wb.create_sheet("Risk_Assessment")

    # Title
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "Transfer Channel Risk Assessment"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Channel", "Threat Scenario", "Likelihood", "Impact", "Inherent Risk", "Current Controls", "Control Effectiveness", "Residual Risk", "Treatment", "Owner"]
    create_header_row(ws, 3, headers)

    # Risk items
    risks = [
        ("Corporate Email", "Data interception in transit", "", "", "", "TLS, S/MIME available", "", "", "", ""),
        ("Corporate Email", "Malware distribution via attachment", "", "", "", "ATP scanning, sandbox", "", "", "", ""),
        ("Corporate Email", "Accidental disclosure to wrong recipient", "", "", "", "DLP, address book", "", "", "", ""),
        ("Cloud Storage", "Unauthorized external sharing", "", "", "", "Sharing policies, DLP", "", "", "", ""),
        ("Cloud Storage", "Data breach at provider", "", "", "", "Encryption at rest", "", "", "", ""),
        ("SFTP Server", "Brute force attack on credentials", "", "", "", "Key-based auth, lockout", "", "", "", ""),
        ("SFTP Server", "Data exfiltration by insider", "", "", "", "Logging, access reviews", "", "", "", ""),
        ("USB/Removable Media", "Loss or theft of media", "", "", "", "Device encryption", "", "", "", ""),
        ("USB/Removable Media", "Malware introduction", "", "", "", "Endpoint protection", "", "", "", ""),
        ("Courier Services", "Package interception", "", "", "", "Tamper-evident packaging", "", "", "", ""),
        ("Courier Services", "Delivery to wrong recipient", "", "", "", "Signature verification", "", "", "", ""),
        ("API Gateway", "API key compromise", "", "", "", "Key rotation, mTLS", "", "", "", ""),
        ("Verbal/Phone", "Eavesdropping", "", "", "", "Secure rooms, awareness", "", "", "", ""),
        ("Fax", "Transmission interception", "", "", "", "Encrypted fax", "", "", "", ""),
    ]

    row = 4
    for risk in risks:
        for col, value in enumerate(risk, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [3, 4, 5, 7, 8, 9, 10]:  # Input fields
                cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"C4:C{row+5}", '"Rare,Unlikely,Possible,Likely,Almost Certain"')
    add_data_validation(ws, f"D4:D{row+5}", '"Negligible,Minor,Moderate,Major,Severe"')
    add_data_validation(ws, f"E4:E{row+5}", '"Low,Medium,High,Critical"')
    add_data_validation(ws, f"G4:G{row+5}", '"Effective,Partially Effective,Ineffective,Not Assessed"')
    add_data_validation(ws, f"H4:H{row+5}", '"Low,Medium,High,Critical"')
    add_data_validation(ws, f"I4:I{row+5}", '"Accept,Mitigate,Transfer,Avoid"')

    set_column_widths(ws, {"A": 18, "B": 35, "C": 15, "D": 12, "E": 14, "F": 25, "G": 18, "H": 14, "I": 12, "J": 18})
    ws.freeze_panes = "A4"

    logger.info("Created Risk_Assessment sheet")


def create_evidence_register_sheet(wb: Workbook):
    """Create the Evidence Register sheet."""
    ws = wb.create_sheet("Evidence_Register")

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Evidence Register - Channel Security Assessment"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Evidence ID", "Evidence Type", "Description", "Related Assessment", "Location/Link", "Date Collected", "Collected By", "Status"]
    create_header_row(ws, 3, headers)

    # Sample entries
    evidence = [
        ("EV-514-CSA-001", "Configuration Export", "Email transport rules export", "Email_Assessment", "", "", "", ""),
        ("EV-514-CSA-002", "Screenshot", "DMARC policy verification", "Email_Assessment", "", "", "", ""),
        ("EV-514-CSA-003", "Audit Report", "Cloud sharing settings review", "Cloud_Services", "", "", "", ""),
        ("EV-514-CSA-004", "Configuration Export", "SFTP server configuration", "File_Transfer", "", "", "", ""),
        ("EV-514-CSA-005", "Policy Document", "USB device management policy", "Physical_Channels", "", "", "", ""),
    ]

    row = 4
    for item in evidence:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [5, 6, 7, 8]:
                cell.fill = INPUT_FILL
        row += 1

    # Add empty rows
    for i in range(10):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"H4:H{row-1}", '"Pending,Collected,Verified,Expired,N/A"')

    set_column_widths(ws, {"A": 18, "B": 20, "C": 35, "D": 20, "E": 30, "F": 15, "G": 18, "H": 12})
    ws.freeze_panes = "A4"

    logger.info("Created Evidence_Register sheet")


def create_approval_signoff_sheet(wb: Workbook):
    """Create the Approval and Sign-Off sheet."""
    ws = wb.create_sheet("Approval_SignOff")

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Assessment Approval and Sign-Off"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Document info
    ws.cell(row=3, column=1, value="Document Information").font = BOLD_FONT
    ws.merge_cells("A3:F3")

    info = [
        ("Document ID:", DOCUMENT_ID, "Version:", "1.0"),
        ("Document Title:", WORKBOOK_NAME, "Date:", GENERATED_DATE),
        ("Control Reference:", CONTROL_ID, "Classification:", "INTERNAL"),
    ]

    row = 4
    for item in info:
        ws.cell(row=row, column=1, value=item[0]).font = BOLD_FONT
        ws.cell(row=row, column=2, value=item[1]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=item[2]).font = BOLD_FONT
        ws.cell(row=row, column=4, value=item[3]).font = NORMAL_FONT
        row += 1

    # Assessment summary
    row += 1
    ws.cell(row=row, column=1, value="Assessment Summary").font = BOLD_FONT
    ws.merge_cells(f"A{row}:F{row}")

    row += 1
    summary_headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A"]
    create_header_row(ws, row, summary_headers)

    row += 1
    areas = ["Email Security", "Cloud Services", "File Transfer", "Physical Channels", "TOTAL"]
    for area in areas:
        ws.cell(row=row, column=1, value=area).font = BOLD_FONT if area == "TOTAL" else NORMAL_FONT
        for col in range(2, 7):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
            cell.alignment = CENTER_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER
        row += 1

    # Approval section
    row += 1
    ws.cell(row=row, column=1, value="Approval Signatures").font = BOLD_FONT
    ws.merge_cells(f"A{row}:F{row}")

    row += 1
    approval_headers = ["Role", "Name", "Signature", "Date", "Status", "Comments"]
    create_header_row(ws, row, approval_headers)

    row += 1
    approvers = [
        ("Assessment Lead", "", "", "", "", ""),
        ("IT Security Manager", "", "", "", "", ""),
        ("Information Security Officer", "", "", "", "", ""),
    ]

    for approver in approvers:
        for col, value in enumerate(approver, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col > 1:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"E{row-3}:E{row-1}", '"Pending,Approved,Rejected,Deferred"')

    set_column_widths(ws, {"A": 25, "B": 20, "C": 20, "D": 15, "E": 15, "F": 30})
    ws.freeze_panes = "A4"

    logger.info("Created Approval_SignOff sheet")


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main():
    """Main execution function."""
    logger.info("=" * 70)
    logger.info(f"{DOCUMENT_ID} {WORKBOOK_NAME} Generator")
    logger.info("=" * 70)

    wb = Workbook()

    create_instructions_sheet(wb)
    create_email_assessment_sheet(wb)
    create_cloud_services_sheet(wb)
    create_file_transfer_sheet(wb)
    create_physical_channels_sheet(wb)
    create_risk_assessment_sheet(wb)
    create_evidence_register_sheet(wb)
    create_approval_signoff_sheet(wb)

    output_path = Path(__file__).parent / OUTPUT_FILENAME
    wb.save(output_path)

    logger.info("=" * 70)
    logger.info(f"SUCCESS: Workbook saved as {OUTPUT_FILENAME}")
    logger.info("=" * 70)

    return 0


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.14 Information Transfer control
# =============================================================================
