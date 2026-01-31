#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
Excel Workbook Sanity Checker - ISMS A.8.6 Capacity Management Dashboard
================================================================================

Domain-specific diagnostic utility for A.8.6 Capacity Management Dashboard
workbook validation.

**Purpose:**
Validates dashboard structure, formulas, data validations, and external links
specific to capacity management dashboard requirements.

**Usage:**
    python3 sanity_check_dashboard_a86.py ISMS-IMP-A.8.6.3_Dashboard_YYYYMMDD.xlsx

**Checks Performed:**
- Sheet structure (13 expected sheets)
- Executive Dashboard KPI structure
- Utilization Summary data sources
- Forecast Summary linkage
- Planning Effectiveness metrics
- Capacity Risks prioritization
- Maturity Assessment scoring
- Formula integrity in calculated fields
- External workbook links (if linking enabled)
- Conditional formatting rules
- Data validation dropdowns
- Evidence Summary completeness

**Exit Codes:**
    0 = All checks passed
    1 = Warnings detected (dashboard usable)
    2 = Critical errors detected (regenerate recommended)

Control Reference: ISO/IEC 27001:2022 Annex A Control A.8.6
Dashboard Type: Executive Compliance Dashboard (Consolidation)
Version: 1.0
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)



def check_a86_dashboard(filename):
    """Domain-specific validation for A.8.6 Capacity Management Dashboard."""
    
    print("=" * 80)
    print("A.8.6 CAPACITY MANAGEMENT DASHBOARD - SANITY CHECK")
    print("=" * 80)
    print(f"File: {filename}\n")
    
    issues = []
    warnings = []
    
    # Load workbook
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"✓ Workbook loaded: {len(wb.sheetnames)} sheets\n")
    except Exception as e:
        print(f"✗ CRITICAL: Cannot load workbook: {e}")
        return 2
    
    # ========================================================================
    # EXPECTED SHEET STRUCTURE
    # ========================================================================
    print("=" * 80)
    print("CHECK 1: SHEET STRUCTURE")
    print("=" * 80)
    
    expected_sheets = [
        "Executive_Dashboard",
        "Capacity_Health_Score",
        "Utilization_Summary",
        "Forecast_Summary",
        "Planning_Effectiveness",
        "Capacity_Risks",
        "Maturity_Assessment",
        "Trend_Charts",
        "Recommendations",
        "Evidence_Summary",
        "Data_Sources",
        "Approval_Sign_Off"
    ]
    
    found_sheets = wb.sheetnames
    missing_sheets = [s for s in expected_sheets if s not in found_sheets]
    extra_sheets = [s for s in found_sheets if s not in expected_sheets]
    
    if missing_sheets:
        for sheet in missing_sheets:
            issues.append(f"  ✗ Missing required sheet: '{sheet}'")
    
    if extra_sheets:
        for sheet in extra_sheets:
            warnings.append(f"  ⚠ Unexpected sheet found: '{sheet}'")
    
    if not missing_sheets and not extra_sheets:
        print("  ✓ All expected sheets present, no extras")
    
    print(f"  Expected: {len(expected_sheets)} sheets")
    print(f"  Found: {len(found_sheets)} sheets")
    
    # ========================================================================
    # EXECUTIVE DASHBOARD KPI VALIDATION
    # ========================================================================
    if "Executive_Dashboard" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 2: EXECUTIVE DASHBOARD KPI STRUCTURE")
        print("=" * 80)
        
        ws = wb["Executive_Dashboard"]
        
        # Check for expected KPI metrics
        expected_kpis = [
            'Capacity Health Score',
            'Resources at Critical',
            'Capacity Exhaustion',
            'Proactive Expansion Ratio',
            'Forecast Accuracy',
            'Capacity Incidents'
        ]
        
        kpis_found = []
        for row in ws.iter_rows(min_row=1, max_row=50, max_col=5):
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value)
                    for kpi in expected_kpis:
                        if kpi in cell_value:
                            if kpi not in kpis_found:
                                kpis_found.append(kpi)
        
        print(f"  KPIs detected: {len(kpis_found)}/{len(expected_kpis)}")
        for kpi in expected_kpis:
            if kpi in kpis_found:
                print(f"  ✓ {kpi} detected")
            else:
                warnings.append(f"  ⚠ KPI not detected: {kpi}")
    
    # ========================================================================
    # UTILIZATION SUMMARY VALIDATION
    # ========================================================================
    if "Utilization_Summary" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 3: UTILIZATION SUMMARY DATA VALIDATION")
        print("=" * 80)
        
        ws = wb["Utilization_Summary"]
        
        expected_validations = {
            'Resource_Type': ['Compute', 'Storage', 'Network', 'Application', 'Cloud'],
            'Threshold_Status': ['OK', 'Warning', 'Critical'],
            'Trend': ['Increasing', 'Stable', 'Decreasing']
        }
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
            
            validation_found = {key: False for key in expected_validations.keys()}
            
            for dv in ws.data_validations.dataValidation:
                if dv.formula1:
                    formula = str(dv.formula1)
                    for key, values in expected_validations.items():
                        if any(v in formula for v in values):
                            validation_found[key] = True
            
            for key, found in validation_found.items():
                if found:
                    print(f"  ✓ {key} validation detected")
                else:
                    warnings.append(f"  ⚠ {key} validation not detected")
        else:
            warnings.append("  ⚠ No data validations found in Utilization Summary")
    
    # ========================================================================
    # FORECAST SUMMARY VALIDATION
    # ========================================================================
    if "Forecast_Summary" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 4: FORECAST SUMMARY DATA VALIDATION")
        print("=" * 80)
        
        ws = wb["Forecast_Summary"]
        
        expected_horizons = ['6 Month', '12 Month', '24 Month']
        
        horizons_found = []
        for row in ws.iter_rows(min_row=1, max_row=20, max_col=10):
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value)
                    for horizon in expected_horizons:
                        if horizon in cell_value and horizon not in horizons_found:
                            horizons_found.append(horizon)
        
        print(f"  Forecast horizons detected: {len(horizons_found)}/{len(expected_horizons)}")
        for horizon in expected_horizons:
            if horizon in horizons_found:
                print(f"  ✓ {horizon} forecast detected")
            else:
                warnings.append(f"  ⚠ {horizon} forecast not detected")
    
    # ========================================================================
    # PLANNING EFFECTIVENESS VALIDATION
    # ========================================================================
    if "Planning_Effectiveness" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 5: PLANNING EFFECTIVENESS METRICS")
        print("=" * 80)
        
        ws = wb["Planning_Effectiveness"]
        
        expected_metrics = [
            'Proactive',
            'Reactive',
            'Forecast Accuracy',
            'MAPE'
        ]
        
        metrics_found = []
        for row in ws.iter_rows(min_row=1, max_row=30, max_col=8):
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value)
                    for metric in expected_metrics:
                        if metric in cell_value and metric not in metrics_found:
                            metrics_found.append(metric)
        
        print(f"  Metrics detected: {len(metrics_found)}/{len(expected_metrics)}")
        for metric in expected_metrics:
            if metric in metrics_found:
                print(f"  ✓ {metric} metric detected")
            else:
                warnings.append(f"  ⚠ {metric} metric not detected")
    
    # ========================================================================
    # CAPACITY RISKS VALIDATION
    # ========================================================================
    if "Capacity_Risks" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 6: CAPACITY RISKS PRIORITIZATION")
        print("=" * 80)
        
        ws = wb["Capacity_Risks"]
        
        expected_risk_categories = ['Critical', 'High', 'Medium', 'Low']
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
            
            risk_found = False
            for dv in ws.data_validations.dataValidation:
                if dv.formula1:
                    formula = str(dv.formula1)
                    if any(risk in formula for risk in expected_risk_categories):
                        risk_found = True
                        print("  ✓ Risk severity validation detected")
                        break
            
            if not risk_found:
                warnings.append("  ⚠ Risk severity validation not detected")
        else:
            warnings.append("  ⚠ No data validations in Capacity Risks")
    
    # ========================================================================
    # MATURITY ASSESSMENT VALIDATION
    # ========================================================================
    if "Maturity_Assessment" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 7: MATURITY ASSESSMENT SCORING")
        print("=" * 80)
        
        ws = wb["Maturity_Assessment"]
        
        expected_dimensions = [
            'Monitoring',
            'Forecasting',
            'Planning',
            'Optimization',
            'Reporting'
        ]
        
        expected_levels = ['Level 1', 'Level 2', 'Level 3', 'Level 4', 'Level 5']
        
        dimensions_found = []
        levels_found = []
        
        for row in ws.iter_rows(min_row=1, max_row=50, max_col=10):
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value)
                    for dim in expected_dimensions:
                        if dim in cell_value and dim not in dimensions_found:
                            dimensions_found.append(dim)
                    for level in expected_levels:
                        if level in cell_value and level not in levels_found:
                            levels_found.append(level)
        
        print(f"  Maturity dimensions detected: {len(dimensions_found)}/{len(expected_dimensions)}")
        for dim in expected_dimensions:
            if dim in dimensions_found:
                print(f"  ✓ {dim} dimension detected")
            else:
                warnings.append(f"  ⚠ {dim} dimension not detected")
        
        if len(levels_found) >= 3:
            print(f"  ✓ Maturity levels detected: {len(levels_found)}/5")
        else:
            warnings.append(f"  ⚠ Insufficient maturity levels: {len(levels_found)}/5")
    
    # ========================================================================
    # FORMULA INTEGRITY CHECK
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 8: FORMULA INTEGRITY")
    print("=" * 80)
    
    formula_count = 0
    formula_errors = 0
    external_links = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
                    formula = cell.value
                    
                    # Check for external links
                    if '[' in formula and ']' in formula:
                        external_links += 1
                    
                    # Check for common errors
                    if formula.count('(') != formula.count(')'):
                        issues.append(f"  ✗ {sheet_name}!{cell.coordinate}: Unbalanced parentheses")
                        formula_errors += 1
                    
                    if formula.count('"') % 2 != 0:
                        issues.append(f"  ✗ {sheet_name}!{cell.coordinate}: Unbalanced quotes")
                        formula_errors += 1
    
    print(f"  Total formulas found: {formula_count}")
    print(f"  External workbook links: {external_links}")
    
    if external_links > 0:
        print("  ℹ️  External linking enabled (expected if using formula approach)")
        print("     Dashboard references Assessment 1 and Assessment 2 workbooks")
        print("     Ensure assessment workbooks are in same directory")
    else:
        print("  ℹ️  No external links (using data consolidation approach)")
    
    if formula_errors == 0:
        print("  ✓ No formula syntax errors detected")
    else:
        print(f"  ✗ Formula errors found: {formula_errors}")
    
    # ========================================================================
    # CONDITIONAL FORMATTING CHECK
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 9: CONDITIONAL FORMATTING")
    print("=" * 80)
    
    total_cf_rules = 0
    sheets_with_cf = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
            cf_count = len(ws.conditional_formatting._cf_rules)
            if cf_count > 0:
                sheets_with_cf += 1
                total_cf_rules += cf_count
    
    print(f"  Total conditional formatting rules: {total_cf_rules}")
    print(f"  Sheets with conditional formatting: {sheets_with_cf}")
    
    if total_cf_rules > 0:
        print("  ✓ Conditional formatting present (status colors, thresholds)")
        if total_cf_rules > 100:
            warnings.append(f"  ⚠ High number of CF rules ({total_cf_rules}) - may impact performance")
    else:
        warnings.append("  ⚠ No conditional formatting detected")
    
    # ========================================================================
    # EVIDENCE SUMMARY STRUCTURE
    # ========================================================================
    if "Evidence_Summary" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 10: EVIDENCE SUMMARY STRUCTURE")
        print("=" * 80)
        
        ws = wb["Evidence_Summary"]
        
        has_content = False
        for row in ws.iter_rows(min_row=1, max_row=15, max_col=10):
            for cell in row:
                if cell.value:
                    has_content = True
                    break
            if has_content:
                break
        
        if has_content:
            print("  ✓ Evidence Summary has content")
        else:
            warnings.append("  ⚠ Evidence Summary appears empty")
        
        # Check for data validations
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
        else:
            warnings.append("  ⚠ No data validations in Evidence Summary")
    
    # ========================================================================
    # DATA SOURCES VALIDATION
    # ========================================================================
    if "Data_Sources" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 11: DATA SOURCES DOCUMENTATION")
        print("=" * 80)
        
        ws = wb["Data_Sources"]
        
        expected_sources = [
            'Assessment 1',
            'Assessment 2',
            'Utilization',
            'Forecast'
        ]
        
        sources_found = []
        for row in ws.iter_rows(min_row=1, max_row=30, max_col=8):
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value)
                    for source in expected_sources:
                        if source in cell_value and source not in sources_found:
                            sources_found.append(source)
        
        print(f"  Data sources documented: {len(sources_found)}/{len(expected_sources)}")
        for source in expected_sources:
            if source in sources_found:
                print(f"  ✓ {source} documented")
            else:
                warnings.append(f"  ⚠ {source} not documented")
    
    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    print("\n" + "=" * 80)
    print("VALIDATION SUMMARY")
    print("=" * 80)
    
    if issues:
        print(f"\n❌ CRITICAL ISSUES: {len(issues)}")
        for issue in issues:
            print(issue)
    
    if warnings:
        print(f"\n⚠️  WARNINGS: {len(warnings)}")
        for warning in warnings:
            print(warning)
    
    if not issues and not warnings:
        print("\n✅ ALL CHECKS PASSED")
        print("\nThe A.8.6 Capacity Management Dashboard appears healthy.")
        print("Ready for executive presentation and compliance reporting.")
    else:
        print("\n" + "=" * 80)
        print("RECOMMENDATIONS:")
        print("=" * 80)
        
        if missing_sheets:
            print("\n• Regenerate dashboard with generate_a86_3_compliance_dashboard.py")
            print("  to ensure all required sheets are present")
        
        if external_links > 0:
            print("\n• External linking detected:")
            print("  - Place Assessment 1 and Assessment 2 in same directory")
            print("  - Open dashboard and click 'Update Links' in Excel")
            print("  - OR use consolidate_a86_dashboard.py for static snapshot")
        
        if warnings:
            print("\n• Review warnings - most are informational")
            print("• Data validations improve data quality but are optional")
            print("• Ensure dashboard KPIs align with organizational requirements")
    
    print("\n" + "=" * 80)
    
    # Return exit code
    if issues:
        return 2
    elif warnings:
        return 1
    else:
        return 0


def main():
    if len(sys.argv) < 2:
        print("=" * 80)
        print("A.8.6 CAPACITY MANAGEMENT DASHBOARD - SANITY CHECK")
        print("=" * 80)
        print("\nUsage: python3 sanity_check_dashboard_a86.py <filename.xlsx>")
        print("\nExample:")
        print("  python3 sanity_check_dashboard_a86.py ISMS-IMP-A.8.6.3_Dashboard_20250128.xlsx")
        print("\nExit codes:")
        print("  0 = All checks passed")
        print("  1 = Warnings detected (dashboard usable)")
        print("  2 = Critical errors detected (regenerate recommended)")
        print("\n" + "=" * 80)
        sys.exit(1)
    
    filename = sys.argv[1]
    exit_code = check_a86_dashboard(filename)
    sys.exit(exit_code)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (syntax validated, structure verified)
# QA_TOOL: Claude Code Deep Scan
# QA_NOTE: STANDARDIZATION - License header, logging, main() pattern applied
# =============================================================================
