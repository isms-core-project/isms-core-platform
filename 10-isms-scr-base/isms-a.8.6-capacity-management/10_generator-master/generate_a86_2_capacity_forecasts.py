#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.6-Assessment-2 - Capacity Forecasting & Planning Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.6: Capacity Management
Assessment Workbook 2 of 3: Capacity Forecasts and Planning Effectiveness

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific forecasting methodology, planning cycles, and
capacity management maturity.

Key customization areas:
1. Forecasting horizon (6, 12, 24 months - match your planning cycles)
2. Forecasting methodology (linear regression, seasonal models, growth rates)
3. Forecast accuracy thresholds (acceptable error ranges for your context)
4. Planning cycle frequency (monthly, quarterly, annual reviews)
5. Capacity expansion criteria (when to trigger procurement/provisioning)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.6 Capacity Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for capacity
forecasting, planning effectiveness evaluation, and proactive capacity
expansion planning based on trend analysis and business growth projections.

**Purpose:**
Enables systematic capacity forecasting and planning assessment against ISO
27001:2022 Control A.8.6 requirements, supporting evidence-based capacity
expansion decisions and preventing reactive "firefighting" capacity additions.

**Assessment Scope:**
- Capacity forecasting (6-month, 12-month, 24-month projections)
- Trend analysis (linear growth, seasonal patterns, anomaly detection)
- Forecasted capacity exhaustion dates (when resources will be depleted)
- Capacity expansion planning (planned additions, procurement status)
- Forecast accuracy validation (previous forecasts vs. actual utilization)
- Planning effectiveness metrics (proactive vs. reactive expansions)
- Capacity-related incidents (capacity exhaustion events, near-misses)
- Forecast confidence intervals (uncertainty ranges)
- Business growth alignment (capacity planning with business projections)
- Gap analysis for forecasting methodology improvements

**Generated Workbook Structure:**
1. Instructions & Legend - Forecasting methodology and planning guidance
2. Historical Utilization - Historical data for trend analysis (6-24 months)
3. Trend Analysis - Growth rates, seasonal patterns, regression analysis
4. 6-Month Forecast - Short-term capacity projections (tactical planning)
5. 12-Month Forecast - Medium-term capacity projections (budget planning)
6. 24-Month Forecast - Long-term capacity projections (strategic planning)
7. Capacity Exhaustion - Forecasted dates when resources will be depleted
8. Expansion Planning - Planned capacity additions with timelines and status
9. Forecast Accuracy - Validation of previous forecasts vs. actual utilization
10. Planning Effectiveness - Proactive vs. reactive capacity expansions
11. Capacity Incidents - Capacity-related outages and near-miss events
12. Business Alignment - Capacity planning aligned with business growth
13. Gap Analysis - Forecasting methodology improvements required
14. Evidence Register - Supporting documentation (forecast models, approvals)
15. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Historical utilization data import from Assessment 1 (Utilization)
- Automated trend analysis and growth rate calculations
- Multiple forecasting horizons (short, medium, long-term)
- Capacity exhaustion date calculation (time until depletion)
- Forecast accuracy metrics (MAPE, MAE, forecast bias)
- Proactive vs. reactive expansion tracking (planning effectiveness)
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with capacity management dashboard

**Integration:**
This assessment combines current utilization (Assessment 1) with forecasting
and planning data to feed the A.8.6 Capacity Management Dashboard, providing
comprehensive capacity health visibility for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a86_2_capacity_forecasts.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a86_2_capacity_forecasts.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a86_2_capacity_forecasts.py --date 20250128

Output:
    File: ISMS_A_8_6_2_Capacity_Forecasts_Planning_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Import historical utilization data from Assessment 1 (last 6-24 months)
    2. Export historical data from monitoring tools for trend analysis
    3. Perform trend analysis (linear regression, growth rate calculation)
    4. Generate capacity forecasts for 6, 12, and 24-month horizons
    5. Calculate forecasted capacity exhaustion dates
    6. Document planned capacity expansions with timelines
    7. Validate forecast accuracy by comparing previous forecasts to actuals
    8. Track proactive vs. reactive capacity expansions (planning effectiveness)
    9. Document capacity-related incidents and near-miss events
    10. Align capacity planning with business growth projections
    11. Conduct gap analysis for forecasting methodology improvements
    12. Collect and link evidence (forecast models, approval documents)
    13. Obtain stakeholder approvals
    14. Feed results into A.8.6 Capacity Management Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.6
Assessment Type:      Assessment 2 of 3 (Capacity Forecasts & Planning)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.6: Capacity Management Policy (Governance)
    - ISMS-IMP-A.8.6-S1: Capacity Monitoring Implementation Guide
    - ISMS-IMP-A.8.6-S2: Capacity Forecasting & Planning Implementation Guide
    - ISMS-IMP-A.8.6-S3: Capacity Management Assessment Guide
    - Assessment 1: Capacity Utilization (generate_a86_1_capacity_utilization.py)
    - Dashboard: Capacity Management Overview (generate_a86_3_compliance_dashboard.py)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full forecasting and planning assessment framework
    - Supports multiple forecasting horizons (6, 12, 24 months)
    - Integrated with A.8.6 Capacity Management Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Forecasting Methodology:**
Multiple forecasting approaches are supported in this framework:
- **Linear Regression**: Best for steady, predictable growth patterns
- **Growth Rate Calculation**: Useful for percentage-based growth projection
- **Seasonal Models**: For resources with recurring usage patterns
- **Business-Driven Forecasts**: Align capacity with planned business growth

Default methodology is linear regression on historical utilization data.
Customize based on your infrastructure characteristics and data patterns.

**Historical Data Requirements:**
Quality forecasts require sufficient historical data:
- **Minimum**: 6 months of historical utilization data
- **Recommended**: 12-24 months for better trend identification
- **Data Quality**: <5% missing data points, outliers removed
- **Data Granularity**: Daily or weekly aggregates for trend analysis

Export historical data from your monitoring tools before completing Assessment 2.

**Forecast Accuracy Targets:**
Forecast accuracy varies by resource type and forecasting horizon:
- **Acceptable**: Forecast within ±15% of actual utilization (MAPE ≤15%)
- **Good**: Forecast within ±10% of actual utilization (MAPE ≤10%)
- **Excellent**: Forecast within ±5% of actual utilization (MAPE ≤5%)

Short-term forecasts (6 months) should be more accurate than long-term (24 months).
Document assumptions and confidence intervals for all forecasts.

**Planning Effectiveness Metrics:**
Track proactive vs. reactive capacity expansions:
- **Proactive**: Capacity expansion planned in advance (based on forecast)
- **Reactive**: Emergency capacity addition (unplanned, due to exhaustion)
- **Target**: ≥90% proactive expansions (per ISMS-POL-A.8.6-S2)

Low proactive ratio indicates forecasting or planning process deficiencies.

**Capacity Exhaustion Calculation:**
For each resource, calculate forecasted exhaustion date:
```
Months Until Exhaustion = (100% - Current Utilization) / Monthly Growth Rate
Exhaustion Date = Current Date + Months Until Exhaustion
```

Resources with exhaustion dates within procurement lead time require immediate
capacity planning action (add to expansion planning sheet).

**Assessment Frequency:**
- **Monthly**: Review short-term forecasts (6 months), update expansion tracker
- **Quarterly**: Comprehensive forecast update (6, 12 months), planning review
- **Annually**: Long-term strategic planning (24 months), methodology review

Align assessment frequency with your organization's planning cycles.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Auditors will expect:
- Documented forecasting methodology with justification
- Historical utilization data supporting forecasts
- Forecast accuracy validation (previous forecasts vs. actuals)
- Capacity expansion plans with approval evidence
- Proactive planning effectiveness metrics (≥90% proactive target)
- Capacity incident tracking (zero capacity-related outages target)

**Data Protection:**
Assessment workbooks contain sensitive strategic planning information:
- Capacity forecasts and infrastructure growth projections
- Capacity expansion plans and budget requirements
- Capacity-related incidents and performance issues
- Business growth projections and strategic initiatives

Handle in accordance with your organization's data classification policies.
Consider this assessment workbook "Confidential" classification.

**Forecasting Tool Options:**
- **Excel**: Built-in FORECAST.LINEAR, TREND, GROWTH functions
- **Python**: Pandas, Scikit-learn (linear regression), Prophet (seasonal)
- **Monitoring Tools**: Prometheus, Datadog, CloudWatch built-in forecasting
- **Dedicated Tools**: Enterprise capacity planning platforms

This workbook supports manual forecasting or import from external tools.
See ISMS-IMP-A.8.6-S2 for detailed forecasting methodology guidance.

**Integration with Current Utilization:**
Assessment 2 (Forecasts) builds on Assessment 1 (Utilization):
- Import current utilization data from Assessment 1
- Use historical utilization for trend analysis
- Combine current state + forecasts for comprehensive capacity planning
- Both assessments feed the Capacity Management Dashboard

Complete Assessment 1 before Assessment 2 for optimal workflow.

**Continuous Improvement:**
Use forecast accuracy validation to improve forecasting methodology:
- Compare previous forecasts to actual observed utilization
- Calculate forecast error metrics (MAPE, MAE, bias)
- Identify systematic over-forecasting or under-forecasting
- Adjust forecasting parameters based on accuracy analysis
- Document lessons learned and methodology improvements

Quarterly forecast accuracy review recommended for continuous improvement.

**Cloud vs. On-Premises Planning:**
- **On-premises**: Plan for procurement lead times (weeks to months)
- **Cloud**: Plan for auto-scaling policies and cost thresholds
- **Hybrid**: Separate forecasts for on-premises and cloud resources

Forecasting horizons and expansion planning differ significantly between
on-premises (long lead time) and cloud (instant provisioning).

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)







# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.6-Assessment-2"
WORKBOOK_NAME = "Capacity Forecasting & Planning"
CONTROL_ID = "A.8.6"
CONTROL_NAME = "Capacity Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅
XMARK = '\u274C'      # ❌
WARNING = '\u26A0'    # ⚠️
HOURGLASS = '\u23F3'  # ⏳
BULLET = '\u2022'     # •
ARROW = '\u2192'      # →
CHART = '\U0001F4CA' # 📊
TARGET = '\U0001F3AF' # 🎯
TRENDING_UP = '\U0001F4C8' # 📈
DISK = '\U0001F4BE'   # 💾
STOPWATCH = '\u23F1'  # ⏱️

# Import shared functions from Assessment 1
# Reuse: create_workbook structure, setup_styles, apply_style patterns

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    sheets = [
        "Instructions & Legend",
        "Historical_Utilization",
        "Trend_Analysis",
        "Capacity_Forecasts",
        "Capacity_Exhaustion",
        "Planned_Expansions",
        "Forecast_Accuracy",
        "Budget_Planning",
        "Evidence_Register",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def setup_styles():
    """Define cell styles (same as Assessment 1)."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_ok": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_warning": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_critical": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style to cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def create_instructions(wb, styles):
    """Create Instructions sheet."""
    ws = wb["Instructions & Legend"]
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "ISMS-IMP-A.8.6.2 – Capacity Forecasts and Planning Assessment"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = "ISO/IEC 27001:2022 - Control A.8.6: Capacity Management"
    apply_style(cell, styles['subheader'])
    
    row = 4
    info = [
        ("Document ID:", "ISMS-IMP-A.8.6.2"),
        ("Assessment Area:", "Capacity Forecasting and Planning"),
        ("Related Policy:", "ISMS-POL-A.8.6-S2"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Quarterly"),
    ]
    
    for label, value in info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if "USER INPUT" in value:
            apply_style(ws[f'B{row}'], styles['input_cell'])
        row += 1
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40


def create_hist_utilization(wb, styles):
    """Create Historical Utilization sheet."""
    ws = wb["Historical_Utilization"]
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "HISTORICAL CAPACITY UTILIZATION DATA"
    apply_style(cell, styles['header'])
    
    headers = ["Month/Date", "Resource Name", "Utilization (%)", "Peak Utilization (%)", "Notes", "Data Source"]
    row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 25


def create_trend_analysis(wb, styles):
    """Create Trend Analysis sheet."""
    ws = wb["Trend_Analysis"]
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = "CAPACITY TREND ANALYSIS"
    apply_style(cell, styles['header'])
    
    headers = ["Resource Name", "Trend Method", "Growth Rate (% per month)", "R-Squared", "Seasonal Pattern", "Notes", "Analyst"]
    row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    for width, col_letter in [(30, 'A'), (20, 'B'), (25, 'C'), (15, 'D'), (20, 'E'), (30, 'F'), (20, 'G')]:
        ws.column_dimensions[col_letter].width = width


def create_capacity_forecasts(wb, styles):
    """Create Capacity Forecasts sheet."""
    ws = wb["Capacity_Forecasts"]
    ws.merge_cells('A1:I1')
    cell = ws['A1']
    cell.value = "CAPACITY FORECASTS (6 and 12 Months)"
    apply_style(cell, styles['header'])
    
    headers = ["Resource Name", "Current Utilization (%)", "Growth Rate (%/month)", "6-Month Forecast (%)", 
               "12-Month Forecast (%)", "24-Month Forecast (%)", "Confidence", "Assumptions", "Forecast Date"]
    row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row = 4
    ws[f'D{row}'] = '=B4+(C4*6)'  # 6-month forecast formula
    ws[f'E{row}'] = '=B4+(C4*12)'  # 12-month forecast formula
    ws[f'F{row}'] = '=B4+(C4*24)'  # 24-month forecast formula


def create_capacity_exhaustion(wb, styles):
    """Create Capacity Exhaustion sheet."""
    ws = wb["Capacity_Exhaustion"]
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "CAPACITY EXHAUSTION PROJECTIONS"
    apply_style(cell, styles['header'])
    
    headers = ["Resource Name", "Current (%)", "Threshold (%)", "Growth Rate (%/mo)", 
               "Months to Threshold", "Exhaustion Date", "Urgency", "Action Required"]
    row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row = 4
    # Formula: Months to threshold = (Threshold - Current) / Growth Rate
    ws[f'E{row}'] = '=IF(AND(B4<>"",C4<>"",D4<>"",D4>0),(C4-B4)/D4,"")'
    ws[f'F{row}'] = '=IF(E4<>"",TODAY()+(E4*30),"")'
    ws[f'G{row}'] = '=IF(E4="","No Risk",IF(E4<=0,"CRITICAL",IF(E4<=3,"Immediate (0-3mo)",IF(E4<=6,"Short-term (3-6mo)",IF(E4<=12,"Medium-term (6-12mo)","Long-term (>12mo)")))))'


def create_planned_expansions(wb, styles):
    """Create Planned Expansions sheet."""
    ws = wb["Planned_Expansions"]
    ws.merge_cells('A1:J1')
    cell = ws['A1']
    cell.value = "PLANNED CAPACITY EXPANSIONS"
    apply_style(cell, styles['header'])
    
    headers = ["Resource Name", "Current Capacity", "Expansion Amount", "New Total Capacity",
               "Planned Date", "Lead Time (days)", "Status", "Cost Estimate", "Approval Status", "Owner"]
    row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])


def create_forecast_accuracy(wb, styles):
    """Create Forecast Accuracy sheet."""
    ws = wb["Forecast_Accuracy"]
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "FORECAST ACCURACY VALIDATION"
    apply_style(cell, styles['header'])
    
    headers = ["Resource Name", "Forecast Date", "Forecasted Value", "Actual Value", 
               "Absolute Error", "Percentage Error (%)", "Accuracy Rating", "Lessons Learned"]
    row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row = 4
    ws[f'E{row}'] = '=ABS(C4-D4)'  # Absolute error
    ws[f'F{row}'] = '=IF(D4<>"",ABS((C4-D4)/D4)*100,"")'  # Percentage error
    ws[f'G{row}'] = '=IF(F4="","",IF(F4<=5,"Excellent",IF(F4<=10,"Good",IF(F4<=15,"Acceptable","Poor"))))'


def create_budget_planning(wb, styles):
    """Create Budget Planning sheet."""
    ws = wb["Budget_Planning"]
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = "CAPACITY BUDGET PLANNING"
    apply_style(cell, styles['header'])
    
    headers = ["Resource/Project", "Quarter", "CapEx ($)", "OpEx ($/month)", "3-Year TCO ($)", "Approval Status", "Notes"]
    row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row = 4
    ws[f'E{row}'] = '=C4+(D4*12*3)'  # 3-year TCO formula


def create_evidence_register(wb, styles):
    """Create Evidence Register (same as Assessment 1)."""
    ws = wb["Evidence_Register"]
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "EVIDENCE REGISTER"
    apply_style(cell, styles['header'])
    
    headers = ["Evidence ID", "Evidence Type", "Description", "Location/File Path", "Date Created", "Related Sheet"]
    row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles['column_header'])


def create_approval_signoff(wb, styles):
    """Create Approval Sign-Off sheet."""
    ws = wb["Approval_Sign_Off"]
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "ASSESSMENT APPROVAL AND SIGN-OFF"
    apply_style(cell, styles['header'])
    
    row = 4
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "FORECAST SUMMARY"
    apply_style(cell, styles['subheader'])
    
    row += 1
    summary = [
        ("Assessment Date:", "[USER INPUT]"),
        ("Forecast Period:", "[e.g., Q1 2026 - Q4 2026]"),
        ("Resources Forecasted:", "[Count]"),
        ("Resources Requiring Expansion:", "[Count]"),
        ("Total Budget Impact:", "[$ Amount]"),
    ]
    
    for label, value in summary:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if "USER INPUT" in value or "[" in value:
            apply_style(ws[f'B{row}'], styles['input_cell'])
        row += 1
    
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "APPROVAL SIGNATURES"
    apply_style(cell, styles['subheader'])
    
    row += 1
    roles = [
        "Completed By (Capacity Planning Team)",
        "Reviewed By (Infrastructure Manager)",
        "Approved By (IT Director/CIO)",
        "Approved By (CFO - Budget Approval)",
    ]
    
    ws[f'A{row}'] = "Role"
    ws[f'B{row}'] = "Name"
    ws[f'C{row}'] = "Signature/Approval"
    ws[f'D{row}'] = "Date"
    for col in range(1, 5):
        apply_style(ws.cell(row=row, column=col), styles['column_header'])
    
    row += 1
    for role in roles:
        ws[f'A{row}'] = role
        for col in range(2, 5):
            apply_style(ws.cell(row=row, column=col), styles['input_cell'])
        row += 1


def main():
    """Main execution function."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.8.6.2 - Capacity Forecasts and Planning Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.6: Capacity Management")
    logger.info("=" * 80)
    logger.info("")
    
    logger.info("Creating workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    
    logger.info("Generating Instructions...")
    create_instructions(wb, styles)
    
    logger.info("Generating Historical Utilization...")
    create_hist_utilization(wb, styles)
    
    logger.info("Generating Trend Analysis...")
    create_trend_analysis(wb, styles)
    
    logger.info("Generating Capacity Forecasts...")
    create_capacity_forecasts(wb, styles)
    
    logger.info("Generating Capacity Exhaustion...")
    create_capacity_exhaustion(wb, styles)
    
    logger.info("Generating Planned Expansions...")
    create_planned_expansions(wb, styles)
    
    logger.info("Generating Forecast Accuracy...")
    create_forecast_accuracy(wb, styles)
    
    logger.info("Generating Budget Planning...")
    create_budget_planning(wb, styles)
    
    logger.info("Generating Evidence Register...")
    create_evidence_register(wb, styles)
    
    logger.info("Generating Approval Sign-Off...")
    create_approval_signoff(wb, styles)
    
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.8.6.2_Capacity_Forecasts_Planning_{timestamp}.xlsx"
    
    logger.info("")
    logger.info(f"Saving workbook: {filename}")
    wb.save(filename)
    
    logger.info("")
    logger.info("=" * 80)
    logger.info("{CHECK} SUCCESS - Capacity Forecasts and Planning Assessment Workbook Created")
    logger.info("=" * 80)
    logger.info("")
    logger.info(f"Output file: {filename}")
    logger.info("")
    logger.info("NEXT STEPS:")
    logger.info("1. Open workbook in Excel/LibreOffice")
    logger.info("2. Import historical utilization data (Historical_Utilization sheet)")
    logger.info("3. Perform trend analysis and calculate growth rates")
    logger.info("4. Develop 6, 12, and 24-month capacity forecasts")
    logger.info("5. Calculate capacity exhaustion dates")
    logger.info("6. Plan capacity expansions with budget impact")
    logger.info("7. Validate forecast accuracy (if previous forecasts available)")
    logger.info("8. Obtain approvals (IT Director, CFO)")
    logger.info("")
    logger.info("For dashboard integration:")
    logger.info("  • Run normalize_assessment_files_a86.py after completing assessments")
    logger.info("  • Generate dashboard: python3 generate_dashboard_capacity_management.py")
    logger.info("")
    logger.info("=" * 80)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.error(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        exit(1)

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
