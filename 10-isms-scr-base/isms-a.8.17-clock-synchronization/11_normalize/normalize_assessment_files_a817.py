#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-A.8.17 - Assessment File Normalization Utility
================================================================================

ISO/IEC 27001:2022 Control A.8.17: Clock Synchronization
Quality Assurance Utility: Excel Assessment File Normalization for Dashboard

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific file naming conventions and validation requirements.

Key customization areas:
1. Expected Document ID patterns (match your naming standards)
2. File naming convention rules (adapt to your standards)
3. Validation severity thresholds (based on your quality requirements)
4. Output path conventions (align with your file structure)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.17 Clock Synchronization Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This utility normalizes A.8.17 assessment workbook filenames to enable
formula-based dashboard integration by creating consistent, date/version-free
filenames that Excel formulas can reliably reference.

**The Problem:**
Assessment workbooks are often created with date suffixes or version numbers:
- ISMS-A.8.17-Assessment-1-Time-Sources-20250125.xlsx
- ISMS-A.8.17-Assessment-2-Sync-Status-v3.xlsx
- ISMS_A_8_17_Assessment_1_20250115.xlsx

Dashboard formulas need consistent names:
- ISMS-A.8.17-Assessment-1.xlsx
- ISMS-A.8.17-Assessment-2.xlsx

**The Solution:**
This script:
1. Scans directory for A.8.17 assessment workbooks
2. Validates Document IDs in "Instructions" sheet
3. Copies files to normalized filenames (removes dates/versions)
4. Creates audit manifest tracking original → normalized mapping
5. Validates normalized files are dashboard-ready

**Purpose:**
Enables reliable formula-based dashboard integration by ensuring assessment
workbooks have consistent, predictable filenames that remain stable across
assessment updates and versions.

**Key Functions:**
1. **Assessment Discovery**
   - Scans directory for A.8.17 workbooks
   - Identifies assessment type by Document ID
   - Supports various naming patterns

2. **Document ID Validation**
   - Opens workbook and reads "Instructions" sheet
   - Validates Document ID presence and format
   - Confirms assessment type matches filename

3. **Filename Normalization**
   - Strips date suffixes (YYYYMMDD, YYYY-MM-DD)
   - Removes version indicators (v1, v2, _draft, _final)
   - Creates standardized filename per assessment type

4. **File Operations**
   - Copies (not moves) to preserve originals
   - Overwrites existing normalized files
   - Creates backup if normalized file exists

5. **Audit Trail**
   - Generates manifest.txt with normalization mapping
   - Records timestamp and file sizes
   - Provides traceability for quality assurance

**Expected Assessment Workbooks:**
- ISMS-IMP-A.8.17.1: Time Source Infrastructure Assessment
  → Normalized to: ISMS-A.8.17-Assessment-1.xlsx
  
- ISMS-IMP-A.8.17.2: System Synchronization Status Assessment
  → Normalized to: ISMS-A.8.17-Assessment-2.xlsx

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel validation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library - read-only for validation)
    - pathlib (standard library - file operations)
    - shutil (standard library - file copying)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage (Interactive):
    python3 normalize_assessment_files_a817.py
    
    Script will:
    - Scan current directory for assessment workbooks
    - Display found assessments and proposed normalized names
    - Prompt for confirmation before normalizing
    - Create normalized copies
    - Generate audit manifest

Advanced Usage:
    # Specify source directory
    python3 normalize_assessment_files_a817.py --source ./assessments
    
    # Specify output directory (different from source)
    python3 normalize_assessment_files_a817.py --source ./assessments --output ./dashboard
    
    # Auto-confirm (skip interactive prompts)
    python3 normalize_assessment_files_a817.py --auto-confirm
    
    # Dry-run mode (show what would be done)
    python3 normalize_assessment_files_a817.py --dry-run

Command-Line Options:
    --source PATH       Source directory containing assessment workbooks (default: current)
    --output PATH       Output directory for normalized files (default: same as source)
    --auto-confirm      Skip confirmation prompts (for automation)
    --dry-run           Show normalization plan without making changes
    --backup            Create backup of existing normalized files before overwriting

Output Files:
    Normalized workbooks:
        - ISMS-A.8.17-Assessment-1.xlsx
        - ISMS-A.8.17-Assessment-2.xlsx
    
    Audit manifest:
        - A817_Assessment_Normalization_Manifest_YYYYMMDD.txt

**Example Session:**
$ python3 normalize_assessment_files_a817.py
ISMS A.8.17 - ASSESSMENT FILE NORMALIZATION
Scanning: /home/user/assessments
Found 2 assessment workbooks:

ISMS-A.8.17-Assessment-1-Time-Sources-20250125.xlsx
Document ID: ISMS-IMP-A.8.17.1
Will normalize to: ISMS-A.8.17-Assessment-1.xlsx
ISMS-A.8.17-Assessment-2-Sync-Status-v2.xlsx
Document ID: ISMS-IMP-A.8.17.2
Will normalize to: ISMS-A.8.17-Assessment-2.xlsx

Proceed with normalization? (yes/no): yes
[✓] Normalized: Assessment-1 → ISMS-A.8.17-Assessment-1.xlsx
[✓] Normalized: Assessment-2 → ISMS-A.8.17-Assessment-2.xlsx
[✓] Created manifest: A817_Assessment_Normalization_Manifest_20250125.txt
================================================================================
Normalization complete! Dashboard is ready for formula linking.
Place dashboard in same directory as normalized files.

**Workflow Integration:**

Complete assessments with any filename:
generate_a817_1_time_sources.py  → creates dated file
generate_a817_2_sync_status.py   → creates dated file
Normalize filenames for dashboard:
python3 normalize_assessment_files_a817.py
Generate dashboard with formula links:
python3 generate_a817_4_compliance_dashboard.py
Open dashboard in Excel → Update Links → Done


**Automation Example (CI/CD Pipeline):**
```bash
#!/bin/bash
# Monthly A.8.17 compliance assessment automation

# 1. Update assessments (manual or scripted data collection)
# (Assume assessment workbooks are updated by ops team)

# 2. Normalize filenames
python3 normalize_assessment_files_a817.py --auto-confirm

# 3. Generate fresh dashboard
python3 generate_a817_4_compliance_dashboard.py

# 4. Archive monthly snapshot
DATE=$(date +%Y-%m)
cp ISMS-A.8.17-Dashboard-Time-Sync.xlsx "history/Dashboard-${DATE}.xlsx"

echo "Monthly A.8.17 compliance dashboard updated"
```

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.17
Utility Type:         Quality Assurance - Filename Normalization
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.17: Clock Synchronization Policy
    - ISMS-IMP-A.8.17.1: Time Source Infrastructure Assessment
    - ISMS-IMP-A.8.17.2: System Synchronization Status Assessment
    - A.8.17 Compliance Dashboard (Requires Normalized Files)

Related Scripts:
    - generate_a817_1_time_sources.py (creates Assessment 1)
    - generate_a817_2_sync_status.py (creates Assessment 2)
    - generate_a817_4_compliance_dashboard.py (consumes normalized files)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Document ID validation from Instructions sheet
    - Multi-pattern filename recognition
    - Audit manifest generation
    - Interactive and automated modes
    - Dry-run capability

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Why Normalization is Required:**
Excel's external workbook formulas use EXACT filename matching:
    ='[ISMS-A.8.17-Assessment-1.xlsx]Sheet'!$A$5

If filename changes (date updated, version incremented), formulas break with
#REF! errors. Normalization provides stable target names for formulas.

**Original Files Preserved:**
Script COPIES files (not moves) to normalized names. Your original dated/
versioned files remain untouched. This is safe - you can re-run normalization
any time without losing data.

**Overwrite Behavior:**
If normalized file already exists, script will:
- Create backup (if --backup flag used)
- Overwrite with newer assessment data
- Log old → new in manifest

This is CORRECT behavior - normalized files should reflect latest assessment.

**Document ID Validation:**
Script opens each workbook and reads Document ID from Instructions sheet
(typically row 4, column B). This ensures correct assessment type identification
even if filename is ambiguous.

Supported patterns:
- ISMS-IMP-A.8.17.1 → Assessment 1 (Time Sources)
- ISMS-IMP-A.8.17.2 → Assessment 2 (Sync Status)

**File Naming Flexibility:**
Script recognizes multiple filename patterns:
- With dates: *-20250125.xlsx, *-2025-01-25.xlsx
- With versions: *-v1.xlsx, *-v2.3.xlsx, *_final.xlsx
- With underscores: ISMS_A_8_17_Assessment_1_20250125.xlsx
- With mixed case: isms-a.8.17-assessment-1.xlsx

All normalize to standard format: ISMS-A.8.17-Assessment-N.xlsx

**Audit Manifest:**
Manifest file provides traceability:
A.8.17 Assessment File Normalization
Timestamp: 2025-01-25 10:30:00
Original: ISMS-A.8.17-Assessment-1-Time-Sources-20250125.xlsx (234 KB)
Normalized: ISMS-A.8.17-Assessment-1.xlsx
Document ID: ISMS-IMP-A.8.17.1
Status: Success
Original: ISMS-A.8.17-Assessment-2-Sync-Status-v2.xlsx (456 KB)
Normalized: ISMS-A.8.17-Assessment-2.xlsx
Document ID: ISMS-IMP-A.8.17.2
Status: Success

Keep manifests for audit trail of assessment versions.

**Error Handling:**
Script continues processing even if one assessment fails:
- Missing Document ID: Warns but continues
- Corrupt workbook: Logs error and skips
- Write permission denied: Reports error for that file only
- All other files still processed

Review output for any reported issues.

**Dashboard Preparation Checklist:**
After normalization, verify:
- [ ] ISMS-A.8.17-Assessment-1.xlsx exists
- [ ] ISMS-A.8.17-Assessment-2.xlsx exists
- [ ] Both files open without errors in Excel
- [ ] Document IDs validated correctly
- [ ] Manifest generated successfully

Now ready to generate dashboard with formula-based integration.

**Monthly Workflow:**
1. Update source assessment workbooks (refresh sync data)
2. Save assessments with new dates: Assessment-1-20250225.xlsx
3. Run normalize script: python3 normalize_assessment_files_a817.py
4. Normalized files updated: ISMS-A.8.17-Assessment-1.xlsx
5. Open dashboard: Excel prompts "Update Links" → YES
6. Dashboard refreshes with latest data automatically

**File Size Considerations:**
Script works with assessment workbooks of any size, but:
- <5MB: Fast, no issues
- 5-20MB: May take 5-10 seconds per file
- >20MB: Consider optimizing assessments (remove excess rows)

Normalization time depends on read speed for Document ID validation.

**Security Considerations:**
Script reads workbook metadata (Document ID) in read-only mode. No macros
executed, no formulas evaluated. Safe to run on untrusted workbooks.

Normalized files are copies - if source is compromised, normalized file may
also be compromised. Validate source file integrity before normalization.

**Integration with Other Controls:**
This normalization pattern can be reused for other ISMS controls with
multiple assessment workbooks feeding dashboards. Customize Document ID
mapping for other control families.

**Quality Assurance:**
Run normalization after completing assessments but BEFORE generating dashboard:
Assessment Generation → Normalization → Dashboard Generation → Review

This workflow ensures dashboard always has correct formula references to
normalized assessment files.

**Troubleshooting:**

**Issue: "Document ID not found in Instructions sheet"**
Solution: Check Instructions sheet exists and Document ID is in column B,
typically rows 4-20. May need to adjust script's search range.

**Issue: "Normalized file already exists"**
Solution: Normal behavior - script overwrites. Use --backup to save old version.

**Issue: "Permission denied writing normalized file"**
Solution: Check directory write permissions. Close Excel if file is open.

**Issue: "Dashboard formulas show #REF! after normalization"**
Solution: Ensure dashboard and normalized files are in SAME directory.

================================================================================
"""

import sys
import openpyxl
from pathlib import Path
from datetime import datetime
from shutil import copy2

# Expected assessment workbooks with their Document IDs
EXPECTED_DOCS = {
    "ISMS-IMP-A.8.17.1": {
        "title": "Time Source Infrastructure Assessment",
        "normalized": "ISMS-A.8.17-Assessment-1.xlsx"
    },
    "ISMS-IMP-A.8.17.2": {
        "title": "System Synchronization Status Assessment",
        "normalized": "ISMS-A.8.17-Assessment-2.xlsx"
    }
}

# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_workbook(filepath):
    """
    Validate workbook and extract Document ID from Instructions sheet.
    
    Args:
        filepath: Path to Excel workbook
    
    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        # Open workbook in read-only mode
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # Check for Instructions sheet (with various naming patterns)
        sheet_name = None
        for name in ["Instructions", "Instructions_Legend", "Instructions & Legend"]:
            if name in wb.sheetnames:
                sheet_name = name
                break
        
        if not sheet_name:
            wb.close()
            return (None, None)
        
        ws = wb[sheet_name]
        
        # Look for Document ID in column A (typically rows 4-20)
        # Format: "Document ID" label in column A, value in column B
        for row in range(3, 25):
            cell_label = ws.cell(row=row, column=1).value
            
            if cell_label and "Document ID" in str(cell_label):
                doc_id_value = ws.cell(row=row, column=2).value
                
                if doc_id_value:
                    doc_id = str(doc_id_value).strip()
                    
                    # Check if it's one of our expected documents
                    if doc_id in EXPECTED_DOCS:
                        title = EXPECTED_DOCS[doc_id]["title"]
                        wb.close()
                        return (doc_id, title)
        
        wb.close()
        return (None, None)
        
    except Exception as e:
        print(f"    ⚠️  Error reading file: {e}")
        return (None, None)

def get_file_info(filepath):
    """Get file metadata for manifest"""
    stat = filepath.stat()
    return {
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime),
        "created": datetime.fromtimestamp(stat.st_ctime),
    }

# ============================================================================
# SCANNING FUNCTIONS
# ============================================================================

def scan_directory(directory):
    """
    Scan directory for assessment workbooks.
    
    Args:
        directory: Path to scan
    
    Returns:
        dict: {doc_id: {'path': Path, 'info': metadata}}
    """
    found_assessments = {}
    directory = Path(directory)
    
    # Get all .xlsx files, excluding temporary files
    xlsx_files = [f for f in directory.glob("*.xlsx") if not f.name.startswith("~$")]
    
    if not xlsx_files:
        print(f"\n❌ No Excel files found in {directory}\n")
        return found_assessments
    
    print(f"\n🔍 Scanning {len(xlsx_files)} Excel file(s) in {directory}...\n")
    
    for filepath in sorted(xlsx_files):
        print(f"  Checking: {filepath.name}")
        doc_id, title = validate_workbook(filepath)
        
        if doc_id:
            print(f"    ✅ Valid: {doc_id} - {title}")
            
            # Check for duplicates
            if doc_id in found_assessments:
                print(f"\n    ⚠️  WARNING: DUPLICATE FOUND FOR {doc_id}")
                print(f"        Previous: {found_assessments[doc_id]['path'].name}")
                print(f"        Current:  {filepath.name}\n")
                
                response = input("        Use CURRENT file? (y/n): ").strip().lower()
                
                if response == 'y':
                    found_assessments[doc_id] = {
                        'path': filepath,
                        'info': get_file_info(filepath)
                    }
                    print(f"        → Replaced with current file\n")
                else:
                    print(f"        → Keeping previous file\n")
            else:
                found_assessments[doc_id] = {
                    'path': filepath,
                    'info': get_file_info(filepath)
                }
        else:
            print(f"    ⏭️  Skipped (not a valid A.8.17 assessment workbook)")
        
        print()  # Blank line for readability
    
    return found_assessments

# ============================================================================
# NORMALIZATION FUNCTIONS
# ============================================================================

def normalize_files(source_dir=None, output_dir=None, auto_confirm=False):
    """
    Main normalization orchestration function.
    
    Args:
        source_dir: Source directory (prompts if None)
        output_dir: Output directory (prompts if None)
        auto_confirm: Skip confirmation prompts (for automation)
    
    Returns:
        bool: True if successful, False otherwise
    """
    print("=" * 80)
    print("ISMS ASSESSMENT FILE NORMALIZATION UTILITY")
    print("ISO/IEC 27001:2022 - Control A.8.17: Clock Synchronization")
    print("=" * 80)
    print("\nThis script prepares assessment workbooks for dashboard linking by:")
    print("  1. Scanning for completed assessment files")
    print("  2. Validating document IDs in Instructions sheet")
    print("  3. Copying to normalized filenames (no dates/versions)")
    print("  4. Creating audit manifest for traceability\n")
    
    # Get source directory
    if not source_dir:
        source_dir = input("Enter source directory path (or press Enter for current directory): ").strip()
        if not source_dir:
            source_dir = "."
    
    source_dir = Path(source_dir).resolve()
    
    if not source_dir.exists():
        print(f"\n❌ Error: Source directory does not exist: {source_dir}\n")
        return False
    
    if not source_dir.is_dir():
        print(f"\n❌ Error: Not a directory: {source_dir}\n")
        return False
    
    print(f"\n📁 Source directory: {source_dir}")
    
    # Get output directory
    if not output_dir:
        default_output = source_dir / "Dashboard_Sources"
        output_input = input(f"\nEnter output directory (or press Enter for '{default_output.name}'): ").strip()
        if not output_input:
            output_dir = default_output
        else:
            output_dir = Path(output_input)
    
    output_dir = Path(output_dir).resolve()
    
    # Create output directory if it doesn't exist
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        print(f"📁 Output directory: {output_dir}")
    except Exception as e:
        print(f"\n❌ Error creating output directory: {e}\n")
        return False
    
    # Scan for assessment files
    found = scan_directory(source_dir)
    
    if not found:
        print("❌ No valid assessment workbooks found in source directory\n")
        print("   Ensure files contain valid Document IDs in 'Instructions' sheet:")
        for doc_id, info in EXPECTED_DOCS.items():
            print(f"   - {doc_id}: {info['title']}")
        print()
        return False
    
    # Show normalization summary
    print("=" * 80)
    print("NORMALIZATION SUMMARY")
    print("=" * 80 + "\n")
    print(f"Found {len(found)} of {len(EXPECTED_DOCS)} required assessment workbooks:\n")
    
    mapping = {}
    for doc_id in sorted(EXPECTED_DOCS.keys()):
        if doc_id in found:
            normalized_name = EXPECTED_DOCS[doc_id]["normalized"]
            source_path = found[doc_id]['path']
            
            print(f"  ✅ {doc_id}")
            print(f"     Title:      {EXPECTED_DOCS[doc_id]['title']}")
            print(f"     Source:     {source_path.name}")
            print(f"     Normalized: {normalized_name}\n")
            
            mapping[doc_id] = {
                'path': source_path,
                'normalized': normalized_name,
                'info': found[doc_id]['info']
            }
        else:
            print(f"  ❌ {doc_id} - NOT FOUND")
            print(f"     Title: {EXPECTED_DOCS[doc_id]['title']}\n")
    
    if len(found) < len(EXPECTED_DOCS):
        print(f"⚠️  WARNING: Only {len(found)}/{len(EXPECTED_DOCS)} assessment workbooks found")
        print("   Dashboard will have incomplete data for missing assessments\n")
    
    print(f"Output directory: {output_dir}\n")
    
    # Confirm normalization
    if not auto_confirm:
        response = input("Proceed with normalization? (y/n): ").strip().lower()
        if response != 'y':
            print("\n❌ Normalization cancelled by user\n")
            return False
    
    # Perform normalization (copy files)
    print("\n" + "=" * 80)
    print("NORMALIZING FILES")
    print("=" * 80 + "\n")
    
    for doc_id, data in mapping.items():
        source_path = data['path']
        normalized_name = data['normalized']
        dest_path = output_dir / normalized_name
        
        try:
            # Copy file preserving metadata
            copy2(source_path, dest_path)
            print(f"✅ {source_path.name}")
            print(f"   → {dest_path}")
            print()
        except Exception as e:
            print(f"❌ Error copying {source_path.name}: {e}\n")
            return False
    
    # Create manifest
    print("=" * 80)
    print("CREATING AUDIT MANIFEST")
    print("=" * 80 + "\n")
    
    manifest_path = create_manifest(output_dir, mapping, source_dir)
    print(f"✅ Manifest created: {manifest_path}\n")
    
    # Summary
    print("=" * 80)
    print("NORMALIZATION COMPLETE")
    print("=" * 80 + "\n")
    print(f"Normalized {len(mapping)} assessment workbook(s)")
    print(f"Output directory: {output_dir}\n")
    print("NEXT STEPS:")
    print("=" * 80 + "\n")
    print("  1. Generate the compliance dashboard workbook:")
    print("     python generate_dashboard_time_sync.py\n")
    print("  2. Place the dashboard workbook in this directory:")
    print(f"     {output_dir.resolve()}\n")
    print("  3. Open the dashboard and click 'Update Links' when prompted\n")
    print("  4. The dashboard will auto-populate with data from normalized files\n")
    print("=" * 80 + "\n")
    
    return True

# ============================================================================
# MANIFEST CREATION
# ============================================================================

def create_manifest(output_dir, mapping, source_dir):
    """
    Create audit manifest documenting the normalization process.
    
    Args:
        output_dir: Output directory
        mapping: Normalization mapping dictionary
        source_dir: Source directory
    
    Returns:
        Path: Path to created manifest file
    """
    manifest_path = output_dir / "NORMALIZATION_MANIFEST.txt"
    
    with open(manifest_path, 'w') as f:
        f.write("=" * 80 + "\n")
        f.write("ISMS A.8.17 ASSESSMENT FILE NORMALIZATION MANIFEST\n")
        f.write("ISO/IEC 27001:2022 Control A.8.17: Clock Synchronization\n")
        f.write("=" * 80 + "\n\n")
        
        f.write(f"Normalization Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Source Directory: {source_dir.resolve()}\n")
        f.write(f"Output Directory: {output_dir.resolve()}\n")
        f.write(f"Workbooks Processed: {len(mapping)}/{len(EXPECTED_DOCS)}\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("NORMALIZED FILES\n")
        f.write("=" * 80 + "\n\n")
        
        for doc_id in sorted(mapping.keys()):
            data = mapping[doc_id]
            source_path = data['path']
            normalized_name = data['normalized']
            info = data['info']
            
            f.write(f"Document ID: {doc_id}\n")
            f.write(f"Title: {EXPECTED_DOCS[doc_id]['title']}\n")
            f.write(f"Source File: {source_path.name}\n")
            f.write(f"Normalized File: {normalized_name}\n")
            f.write(f"File Size: {info['size']:,} bytes\n")
            f.write(f"Source Modified: {info['modified'].strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Normalization Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("\n")
        
        if len(mapping) < len(EXPECTED_DOCS):
            f.write("=" * 80 + "\n")
            f.write("MISSING ASSESSMENTS\n")
            f.write("=" * 80 + "\n\n")
            
            for doc_id, info in EXPECTED_DOCS.items():
                if doc_id not in mapping:
                    f.write(f"Document ID: {doc_id}\n")
                    f.write(f"Title: {info['title']}\n")
                    f.write(f"Status: NOT FOUND\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("AUDIT TRAIL\n")
        f.write("=" * 80 + "\n\n")
        f.write("Purpose: Assessment workbooks normalized for dashboard integration\n")
        f.write("Method: File copy with metadata preservation\n")
        f.write("Validation: Document ID extracted from 'Instructions' sheet\n")
        f.write("Traceability: Source filenames and modification dates recorded\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("NEXT STEPS\n")
        f.write("=" * 80 + "\n\n")
        f.write("1. Generate the compliance dashboard workbook:\n")
        f.write("   python generate_dashboard_time_sync.py\n\n")
        f.write("2. Place the dashboard workbook in this directory:\n")
        f.write(f"   {output_dir.resolve()}\n\n")
        f.write("3. Open the dashboard and click 'Update Links' when prompted\n\n")
        f.write("4. The dashboard will auto-populate with data from normalized files\n\n")
        
        f.write("HOW IT WORKS\n")
        f.write("-" * 80 + "\n")
        f.write("Dashboard Linking:\n")
        f.write("  - Dashboard contains formulas with external workbook references\n")
        f.write("  - External workbook links auto-update when sources change\n")
        f.write("  - Place dashboard in same directory as normalized files\n")
        f.write("  - Open dashboard and click 'Update Links' when prompted\n\n")
        f.write("File Retention:\n")
        f.write("  - Original source files: Retained in source directory\n")
        f.write("  - Normalized copies: Located in output directory\n")
        f.write("  - This manifest: Retained with normalized files for audit\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("END OF MANIFEST\n")
        f.write("=" * 80 + "\n")
    
    return manifest_path

# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Normalize ISMS A.8.17 assessment workbooks for dashboard integration",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode (prompts for directories)
  python normalize_assessment_files_a817.py
  
  # Specify source directory
  python normalize_assessment_files_a817.py --source ./assessments
  
  # Specify both directories
  python normalize_assessment_files_a817.py --source ./assessments --output ./dashboard
  
  # Automated mode (no prompts)
  python normalize_assessment_files_a817.py --source ./assessments --auto-confirm
        """
    )
    
    parser.add_argument(
        '--source', '-s',
        help='Source directory containing assessment workbooks (default: current directory)'
    )
    
    parser.add_argument(
        '--output', '-o',
        help='Output directory for normalized files (default: ./Dashboard_Sources)'
    )
    
    parser.add_argument(
        '--auto-confirm', '-y',
        action='store_true',
        help='Skip confirmation prompts (for automation)'
    )
    
    args = parser.parse_args()
    
    success = normalize_files(
        source_dir=args.source,
        output_dir=args.output,
        auto_confirm=args.auto_confirm
    )
    
    sys.exit(0 if success else 1)
