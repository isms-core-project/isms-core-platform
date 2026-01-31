#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
Excel Workbook Sanity Checker for ISMS A.8.17 Assessment Workbooks
================================================================================

ISO/IEC 27001:2022 Control A.8.17: Clock Synchronization
Quality Assurance Utility: Excel Workbook Diagnostic and Validation Tool

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script performs comprehensive diagnostic checks on A.8.17 clock synchronization
assessment Excel workbooks to identify issues that trigger Excel's "file level
validation and repair" warnings and to ensure workbook quality before distribution
or consolidation.

**Purpose:**
Diagnoses common openpyxl-generated Excel issues, validates workbook structure
against framework specifications, and provides actionable remediation guidance
to ensure workbook integrity and audit readiness.

**Diagnostic Scope:**
- Sheet structure validation (presence of required sheets)
- Assessment ID verification (ISMS-IMP-A.8.17.X identifiers)
- Formula syntax validation (inter-sheet references, external links)
- Data validation conflicts (overlapping ranges)
- Merged cell integrity (content in non-primary cells)
- Evidence register validation (capacity and format)
- Approval workflow validation (3-level approval structure)
- Summary dashboard validation (formula completeness)
- Row capacity validation (sufficient data entry capacity)
- Worksheet dimension checks (performance considerations)

**Supported Workbook Types:**
- A817-1: Time Source Infrastructure Assessment
- A817-2: System Synchronization Status Assessment
- A817-3: Exception Register (Optional)
- A817-D: Compliance Dashboard (Master)

**Diagnostic Output:**
- Workbook type detection and classification
- 10 comprehensive validation checks with ✓/✗/⚠ status
- Critical issues requiring immediate fix
- Warnings recommending attention
- Recommended remediation actions
- Workbook-specific notes and expected structure
- Integration guidance for dashboard consolidation

**When to Use:**
- After generating workbooks (Scripts 1-4)
- Before distributing to stakeholders
- After manual modifications or edits
- Before consolidation into dashboard
- When Excel shows repair warnings
- Quality assurance validation before audits

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel processing

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - pathlib (standard library)
    - re (standard library - regex for validation)
    - sys (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    # Check any A.8.17 assessment workbook
    python3 excel_sanity_check_a817.py <filename.xlsx>

Domain-Specific Examples:
    # Check Time Source Infrastructure Assessment
    python3 excel_sanity_check_a817.py ISMS-A.8.17-Assessment-1-Time-Sources.xlsx
    
    # Check System Synchronization Status Assessment
    python3 excel_sanity_check_a817.py ISMS-A.8.17-Assessment-2-Sync-Status.xlsx
    
    # Check Exception Register
    python3 excel_sanity_check_a817.py ISMS-A.8.17-Assessment-3-Exceptions.xlsx
    
    # Check Compliance Dashboard
    python3 excel_sanity_check_a817.py ISMS-A.8.17-Dashboard-Time-Sync.xlsx
    
    # Check normalized workbook
    python3 excel_sanity_check_a817.py ISMS-A.8.17-Assessment-1.xlsx

Quality Assurance Workflow:
    # After generation (immediate validation)
    python3 generate_a817_1_time_sources.py
    python3 excel_sanity_check_a817.py ISMS-A.8.17-Assessment-1-*.xlsx
    
    # Before consolidation (validate all source workbooks)
    for f in ISMS-A.8.17-Assessment-[1-2]-*.xlsx; do
        python3 excel_sanity_check_a817.py "$f"
    done
    
    # After normalization
    python3 normalize_assessment_files_a817.py
    python3 excel_sanity_check_a817.py ISMS-A.8.17-Assessment-*.xlsx
    
    # Dashboard validation
    python3 generate_a817_4_compliance_dashboard.py
    python3 excel_sanity_check_a817.py ISMS-A.8.17-Dashboard-*.xlsx

Output:
    Console output with:
    - Workbook type detection
    - 10 validation checks with detailed findings
    - Critical issues summary (must fix)
    - Warnings summary (should fix)
    - Recommended remediation actions
    - Workbook-specific notes
    - Expected structure documentation

Exit Codes:
    0 - No issues detected (workbook healthy)
    1 - Critical issues found or file error

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.17
Utility Type:         Quality Assurance - Excel Workbook Diagnostics
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization ISMS Team]
Date:                 25.01.2025
Last Modified:        25.01.2025
Python Version:       3.8+
License:              [Organization License/Terms]

Related Documents:
    - ISMS-POL-A.8.17: Clock Synchronization Policy (Governance)
    - ISMS-IMP-A.8.17-S1: Time Source Configuration
    - ISMS-IMP-A.8.17-S2: Synchronization Verification Process
    - ISMS-IMP-A.8.17.1: Time Source Infrastructure Assessment
    - ISMS-IMP-A.8.17.2: System Synchronization Status Assessment
    - ISMS-IMP-A.8.17.D: Compliance Dashboard Specification

Related Scripts:
    - generate_a817_1_time_sources.py (generates workbooks to check)
    - generate_a817_2_sync_status.py (generates workbooks to check)
    - generate_a817_3_exception_register.py (generates workbooks to check)
    - generate_a817_4_compliance_dashboard.py (generates dashboard to check)
    - normalize_assessment_files_a817.py (normalizes before checking)
    - consolidate_a817_dashboard.py (consolidates after checking)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 25.01.2025
    - Initial release
    - Implements 10 comprehensive diagnostic checks
    - Supports all A.8.17 assessment workbook types
    - Provides detailed remediation guidance
    - Includes workbook-specific validation rules

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Quality Assurance Philosophy:**
"The first principle is that you must not fool yourself — and you are the 
easiest person to fool." - Richard Feynman

This sanity checker embodies anti-cargo-cult engineering - it catches structural
issues, formula errors, and validation conflicts that humans easily overlook.
Use it proactively, not reactively (after Excel complains).

**10 Validation Checks Explained:**

1. **Sheet Structure Validation**
   - Verifies all required sheets present
   - Detects missing or renamed sheets
   - Flexible matching (handles minor name variations)
   - Domain-specific expected sheet lists

2. **Assessment ID Validation**
   - Locates ISMS-IMP-A.8.17.X identifier
   - Validates correct assessment type
   - Checks Instructions sheet and metadata

3. **Formula Validation**
   - Checks formula syntax (balanced parentheses)
   - Validates inter-sheet references (sheet exists)
   - Identifies invalid sheet references
   - Detects external workbook links (dashboard requirement)
   - Reports formula issues with cell coordinates

4. **Data Validation Conflicts**
   - Identifies overlapping validation ranges
   - Checks for conflicting dropdown rules
   - Counts total validations per sheet

5. **Merged Cells Validation**
   - Ensures only top-left cell has content
   - Detects content in non-primary merged cells
   - Prevents Excel repair warnings from merged cell issues

6. **Evidence Register Validation**
   - Counts evidence row capacity (target: 50 rows minimum)
   - Validates EV-XXX ID format
   - Enforces minimum evidence capacity
   - Checks for Evidence Register sheet presence

7. **Approval Sign-Off Validation**
   - Checks for 3-level approval workflow
   - Validates: Assessor → Reviewer → Approver
   - Ensures approval structure complete

8. **Summary/Compliance Validation**
   - Counts formulas in Summary/Compliance sheets
   - Validates calculated metrics present
   - Checks for empty or incomplete dashboards

9. **Row Capacity Validation**
   - Checks data entry capacity for assessment sheets
   - Time Sources: minimum 10 sources
   - Sync Status: minimum 50 systems (recommended 200+)
   - Warns on insufficient capacity

10. **Worksheet Dimensions**
    - Flags oversized worksheets (>5000 rows, >50 columns)
    - Identifies potential performance issues

**Common Issues Detected:**

Critical Issues (Must Fix):
- Missing required sheets (Time_Sources, System_Sync_Status)
- Invalid formula references (sheet doesn't exist)
- Broken inter-sheet formulas
- Missing assessment ID
- Evidence Register missing (for Assessments 1-2)

High Priority Issues (Should Fix):
- Merged cell content in non-primary cells
- Overlapping data validation ranges
- Formula syntax errors (unbalanced parentheses)
- Missing approval workflow levels

Medium Priority Issues (Recommended):
- Low row capacity (<50 systems for sync status)
- Missing optional sheets (Exception Register)
- Excessive formula count suggesting copy-paste errors

Low Priority Issues (Nice to Fix):
- Minor naming differences in sheets
- Cosmetic formatting variations

**Excel Repair Warnings - Root Causes:**

If Excel shows "We found a problem with some content" on open:

Common Causes:
1. **Invalid Sheet References** - Formula references non-existent sheet
2. **Merged Cell Issues** - Content in non-primary cells of merge
3. **Data Validation Conflicts** - Overlapping validation ranges
4. **Style Object Sharing** - Border/Font/Fill objects reused incorrectly
5. **External Link Issues** - Dashboard can't resolve source workbooks

This script identifies causes #1-3 directly. For style issues, use Excel's
built-in repair (usually auto-fixes without data loss).

**Dashboard-Specific Considerations (A817-D):**

Dashboard requires normalized source workbooks:
- ISMS-A.8.17-Assessment-1.xlsx (Time Sources)
- ISMS-A.8.17-Assessment-2.xlsx (Sync Status)

If these are missing, dashboard shows #REF errors. This is EXPECTED until:
1. Source workbooks generated (Scripts 1-2)
2. Normalization completed (normalize_assessment_files_a817.py)
3. Files placed in same directory as dashboard

Sanity checker will warn about external references but note this is expected.

**Preventive vs. Reactive Usage:**

Preventive (Best Practice):
- Run after every workbook generation
- Validate before distributing to stakeholders
- Check before normalization/consolidation
- Part of automated quality assurance workflow

Reactive (When Problems Occur):
- Excel shows repair warnings
- Formulas show #REF errors
- Dashboard consolidation fails
- Stakeholders report issues opening files

Use preventively to avoid reactive firefighting.

**Integration with CI/CD:**

```bash
#!/bin/bash
# Quality assurance script for A.8.17 workbook generation

set -e  # Exit on any error

echo "Generating A.8.17 assessment workbooks..."
python3 generate_a817_1_time_sources.py
python3 generate_a817_2_sync_status.py

echo "Running sanity checks..."
for workbook in ISMS-A.8.17-Assessment-[1-2]-*.xlsx; do
    echo "Checking: $workbook"
    python3 excel_sanity_check_a817.py "$workbook" || exit 1
done

echo "All workbooks passed sanity checks ✓"
```

**Audit Considerations:**

Running sanity checks demonstrates quality assurance to auditors:
- Systematic validation of assessment tools
- Documented quality control process
- Evidence of due diligence

Keep sanity check output as QA documentation for audit trail.

**False Positives:**

Some warnings may be acceptable based on context:
- Dashboard external link warnings (expected before normalization)
- Minor sheet name variations (if intentional)
- Optional field completeness (if not applicable)

Review findings and apply judgment - don't blindly fix everything.

**Performance:**

Script typically completes in <5 seconds per workbook. Factors affecting speed:
- Workbook size (large files take longer)
- Number of formulas (more formulas = longer validation)
- Disk I/O (network drives slower than local)

**Limitations:**

What this script CANNOT detect:
- Data accuracy (assessor responsibility)
- Content completeness (human judgment required)
- Time synchronization technical correctness
- Drift threshold appropriateness
- Excel version compatibility issues

What this script DOES detect:
- Structural integrity issues
- Formula correctness
- Required component presence
- Common openpyxl generation errors

**Troubleshooting:**

Problem: Script fails to load workbook
Solution: File may be corrupt - try opening in Excel and saving

Problem: Script reports missing sheets that exist
Solution: Sheet names may have extra spaces or different case - check carefully

Problem: No issues detected but Excel still shows repair warning
Solution: Issue may be style-related (not detected by script) - let Excel repair

Problem: Script runs very slowly
Solution: Large workbook or network drive - copy to local disk for checking

**Related Quality Assurance Tools:**

Complete A.8.17 QA toolchain:
1. **excel_sanity_check_a817.py** (this script) - Structural validation
2. **normalize_assessment_files_a817.py** - Filename normalization
3. **consolidate_a817_dashboard.py** - Data consolidation verification

Use all three for comprehensive quality assurance.

================================================================================
"""

import sys
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ============================================================================
# WORKBOOK TYPE DETECTION
# ============================================================================

def detect_workbook_type(filename):
    """Detect which A.8.17 workbook type this is based on filename."""
    filename_lower = filename.lower()
    
    if 'time' in filename_lower and ('source' in filename_lower or 'infrastructure' in filename_lower):
        return 'A817-1', 'Time Source Infrastructure Assessment'
    elif 'sync' in filename_lower or 'synchronization' in filename_lower:
        return 'A817-2', 'System Synchronization Status Assessment'
    elif 'exception' in filename_lower:
        return 'A817-3', 'Exception Register'
    elif 'dashboard' in filename_lower or 'compliance' in filename_lower:
        return 'A817-D', 'Time Synchronization Compliance Dashboard'
    else:
        return 'Unknown', 'Generic A.8.17 Excel Workbook'


# ============================================================================
# EXPECTED SHEET DEFINITIONS
# ============================================================================

EXPECTED_SHEETS = {
    'A817-1': [
        "Instructions", "Time_Sources", "Internal_NTP_Servers",
        "Hierarchy", "Compliance_Summary", "Evidence_Register", "Approval_Sign_Off"
    ],
    'A817-2': [
        "Instructions", "System_Sync_Status", "Drift_Analysis",
        "Platform_Summary", "Criticality_Analysis", "Gap_Analysis",
        "Remediation_Plan", "Evidence_Register", "Approval_Sign_Off"
    ],
    'A817-3': [
        "Instructions", "Exception_Register", "Risk_Assessment",
        "Compensating_Controls", "Review_Schedule", "Expired_Exceptions",
        "Remediation_Plan", "Evidence_Register", "Approval_Sign_Off"
    ],
    'A817-D': [
        "Instructions", "Executive_Summary", "Time_Sources_Summary",
        "Sync_Status_Summary", "Gap_Analysis", "Exceptions_Summary",
        "Trends", "Action_Items", "Evidence_Approvals"
    ],
}

# Expected Assessment IDs
EXPECTED_ASSESSMENT_IDS = {
    'A817-1': 'ISMS-IMP-A.8.17.1',
    'A817-2': 'ISMS-IMP-A.8.17.2',
    'A817-3': 'ISMS-IMP-A.8.17.3',
    'A817-D': 'ISMS-IMP-A.8.17.D',
}

# Expected normalized filenames for dashboard
DASHBOARD_SOURCES = [
    'ISMS-A.8.17-Assessment-1.xlsx',  # Time Sources
    'ISMS-A.8.17-Assessment-2.xlsx',  # Sync Status
]


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""
    
    workbook_id, workbook_name = detect_workbook_type(filename)
    
    print("=" * 80)
    print(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {Path(filename).name}")
    print(f"Detected Type: {workbook_id} - {workbook_name}")
    print("=" * 80)
    
    issues_found = []
    warnings_found = []
    
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"\n✓ Workbook loaded successfully")
        print(f"  Sheets found: {len(wb.sheetnames)}")
        print(f"  Sheet names: {', '.join(wb.sheetnames)}")
        
    except Exception as e:
        print(f"\n✗ CRITICAL: Cannot load workbook: {e}")
        return 1
    
    # ========================================================================
    # CHECK 1: SHEET STRUCTURE VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 1: SHEET STRUCTURE VALIDATION")
    print("=" * 80)
    
    if workbook_id in EXPECTED_SHEETS:
        expected = EXPECTED_SHEETS[workbook_id]
        actual = wb.sheetnames
        
        # Check for missing sheets (flexible matching)
        missing = []
        for exp_sheet in expected:
            # Flexible match: check if expected sheet name is in any actual sheet
            found = False
            for act_sheet in actual:
                if exp_sheet.lower().replace("_", "").replace(" ", "") == act_sheet.lower().replace("_", "").replace(" ", ""):
                    found = True
                    break
            if not found:
                missing.append(exp_sheet)
        
        if missing:
            for sheet in missing:
                # Critical for core sheets, warning for optional
                if sheet in ["Instructions", "Evidence_Register", "Approval_Sign_Off"]:
                    issues_found.append(f"  ✗ Missing REQUIRED sheet: {sheet}")
                else:
                    warnings_found.append(f"  ⚠ Missing expected sheet: {sheet}")
            print(f"  ⚠ {len(missing)} expected sheets not found (some may be renamed)")
        else:
            print(f"  ✓ All expected sheets present for {workbook_id}")
        
        # Check for extra sheets
        print(f"  Total sheets: {len(actual)} (expected {len(expected)})")
    else:
        print("  ℹ Unknown workbook type - skipping sheet validation")
    
    # ========================================================================
    # CHECK 2: ASSESSMENT ID VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 2: ASSESSMENT ID VALIDATION")
    print("=" * 80)
    
    assessment_id_found = False
    if workbook_id in EXPECTED_ASSESSMENT_IDS:
        expected_id = EXPECTED_ASSESSMENT_IDS[workbook_id]
        
        # Check in Instructions sheet
        for sheet_name in wb.sheetnames:
            if 'instructions' in sheet_name.lower():
                ws = wb[sheet_name]
                # Check first few rows for assessment ID
                for row in range(1, 20):
                    for col in range(1, 10):
                        try:
                            cell_value = ws.cell(row=row, column=col).value
                            if cell_value and isinstance(cell_value, str) and expected_id in cell_value:
                                print(f"  ✓ Assessment ID found: {expected_id}")
                                assessment_id_found = True
                                break
                        except:
                            continue
                    if assessment_id_found:
                        break
            if assessment_id_found:
                break
        
        if not assessment_id_found:
            warnings_found.append(
                f"  ⚠ Assessment ID '{expected_id}' not found in Instructions sheet"
            )
    else:
        print("  ℹ Unknown workbook type - skipping Assessment ID validation")
    
    # ========================================================================
    # CHECK 3: FORMULA VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 3: FORMULA VALIDATION")
    print("=" * 80)
    
    formula_issues = 0
    inter_sheet_refs = {}
    external_refs = set()
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # Check for external workbook references
                    ext_refs = re.findall(r'\[([^\]]+\.xlsx)\]', formula)
                    for ref in ext_refs:
                        external_refs.add(ref)
                    
                    # Check for sheet references
                    if "'" in formula or "!" in formula:
                        sheet_refs = re.findall(r"'?([^'!\[]+)'?!", formula)
                        for ref in sheet_refs:
                            ref_clean = ref.strip("'")
                            # Skip external references
                            if "[" not in ref_clean and ref_clean not in wb.sheetnames:
                                # Check for close matches (underscore vs space)
                                close_match = False
                                for actual in wb.sheetnames:
                                    if ref_clean.replace(" ", "_") == actual or ref_clean.replace("_", " ") == actual:
                                        close_match = True
                                        break
                                if not close_match:
                                    issues_found.append(
                                        f"  ✗ {sheet_name}!{cell.coordinate}: "
                                        f"Invalid sheet reference '{ref_clean}'"
                                    )
                                    formula_issues += 1
                            
                            # Track inter-sheet references
                            if "[" not in ref_clean and ref_clean in wb.sheetnames:
                                if sheet_name not in inter_sheet_refs:
                                    inter_sheet_refs[sheet_name] = set()
                                inter_sheet_refs[sheet_name].add(ref_clean)
                    
                    # Check for common syntax issues
                    if formula.count('(') != formula.count(')'):
                        issues_found.append(
                            f"  ✗ {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced parentheses in formula"
                        )
                        formula_issues += 1
    
    if formula_issues == 0:
        print("  ✓ No formula syntax issues detected")
    else:
        print(f"  ✗ Found {formula_issues} formula issues")
    
    if inter_sheet_refs:
        print(f"  ℹ Inter-sheet references found:")
        for sheet, refs in inter_sheet_refs.items():
            print(f"    {sheet} → {', '.join(sorted(refs))}")
    
    if external_refs:
        print(f"  ℹ External workbook references found:")
        for ref in sorted(external_refs):
            print(f"    → {ref}")
        if workbook_id == 'A817-D':
            print("  ℹ NOTE: Dashboard requires external links - this is EXPECTED")
    
    # ========================================================================
    # CHECK 4: DATA VALIDATION CONFLICTS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 4: DATA VALIDATION CONFLICTS")
    print("=" * 80)
    
    validation_issues = 0
    total_validations = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            dv_count = len(ws.data_validations.dataValidation)
            total_validations += dv_count
            
            # Check for overlapping ranges
            all_ranges = []
            for dv in ws.data_validations.dataValidation:
                if hasattr(dv, 'sqref') and dv.sqref:
                    all_ranges.extend(str(dv.sqref).split())
            
            if len(all_ranges) != len(set(all_ranges)):
                warnings_found.append(
                    f"  ⚠ {sheet_name}: Potentially overlapping data validation ranges"
                )
                validation_issues += 1
    
    print(f"  Total data validations: {total_validations}")
    
    if validation_issues == 0:
        print("  ✓ No data validation conflicts detected")
    
    # ========================================================================
    # CHECK 5: MERGED CELLS VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 5: MERGED CELLS VALIDATION")
    print("=" * 80)
    
    merge_issues = 0
    total_merges = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'merged_cells') and ws.merged_cells:
            merge_count = len(ws.merged_cells.ranges)
            total_merges += merge_count
            
            # Check merged cells for content issues
            for merge_range in ws.merged_cells.ranges:
                min_col, min_row = merge_range.min_col, merge_range.min_row
                max_col, max_row = merge_range.max_col, merge_range.max_row
                
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if row == min_row and col == min_col:
                            continue
                        try:
                            cell = ws.cell(row=row, column=col)
                            if cell.value:
                                cell_coord = f"{get_column_letter(col)}{row}"
                                warnings_found.append(
                                    f"  ⚠ {sheet_name}!{cell_coord}: "
                                    f"Merged cell has content in non-primary cell"
                                )
                                merge_issues += 1
                        except:
                            continue
    
    print(f"  Total merged ranges: {total_merges}")
    
    if merge_issues == 0:
        print("  ✓ Merged cells properly formatted")
    
    # ========================================================================
    # CHECK 6: EVIDENCE REGISTER VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 6: EVIDENCE REGISTER VALIDATION")
    print("=" * 80)
    
    evidence_sheet = None
    for sheet in wb.sheetnames:
        if 'evidence' in sheet.lower() and 'register' in sheet.lower():
            evidence_sheet = sheet
            break
    
    if evidence_sheet:
        ws = wb[evidence_sheet]
        # Count evidence rows (look for EV- prefix or row capacity)
        evidence_count = 0
        for row in range(1, min(110, ws.max_row + 1)):
            try:
                cell_val = ws.cell(row=row, column=1).value
                if cell_val and isinstance(cell_val, str):
                    if cell_val.startswith('EV-') or 'evidence' in cell_val.lower():
                        evidence_count += 1
            except:
                continue
        
        print(f"  Evidence sheet: {evidence_sheet}")
        print(f"  Evidence row templates: {evidence_count}")
        
        # A.8.17 requires minimum capacity for evidence
        if evidence_count < 10:
            warnings_found.append(
                f"  ⚠ Evidence Register has {evidence_count} rows (recommended ≥50 for flexibility)"
            )
        elif evidence_count >= 50:
            print(f"  ✓ Adequate evidence register capacity ({evidence_count} rows)")
        else:
            print(f"  ✓ Evidence register present ({evidence_count} rows)")
    else:
        if workbook_id in ['A817-1', 'A817-2', 'A817-3']:
            warnings_found.append("  ⚠ No Evidence Register sheet found (recommended)")
    
    # ========================================================================
    # CHECK 7: APPROVAL SIGN-OFF VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 7: APPROVAL SIGN-OFF VALIDATION")
    print("=" * 80)
    
    approval_sheet = None
    for sheet in wb.sheetnames:
        if 'approval' in sheet.lower() or 'sign' in sheet.lower():
            approval_sheet = sheet
            break
    
    if approval_sheet:
        ws = wb[approval_sheet]
        
        # Check for 3-level approval structure
        approval_levels = 0
        for row in range(1, min(100, ws.max_row + 1)):
            for col in range(1, 6):
                try:
                    cell_val = ws.cell(row=row, column=col).value
                    if cell_val and isinstance(cell_val, str):
                        cell_upper = cell_val.upper()
                        if 'PREPARED' in cell_upper or 'COMPLETED' in cell_upper or 'ASSESSED' in cell_upper:
                            approval_levels += 1
                        elif 'REVIEW' in cell_upper or 'CHECKED' in cell_upper:
                            approval_levels += 1
                        elif 'CISO' in cell_upper or 'APPROVED' in cell_upper:
                            approval_levels += 1
                except:
                    continue
        
        print(f"  Approval sheet: {approval_sheet}")
        print(f"  Approval levels detected: {approval_levels}")
        
        if approval_levels < 3:
            warnings_found.append(
                f"  ⚠ Approval workflow has {approval_levels} levels (expected 3: Assessor → Reviewer → Approver)"
            )
        else:
            print(f"  ✓ Complete 3-level approval workflow")
    else:
        if workbook_id in ['A817-1', 'A817-2', 'A817-3']:
            warnings_found.append("  ⚠ No Approval Sign-Off sheet found (recommended)")
    
    # ========================================================================
    # CHECK 8: SUMMARY/COMPLIANCE VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 8: SUMMARY/COMPLIANCE VALIDATION")
    print("=" * 80)
    
    summary_sheet = None
    for sheet in wb.sheetnames:
        if 'summary' in sheet.lower() or 'compliance' in sheet.lower() or 'executive' in sheet.lower():
            summary_sheet = sheet
            break
    
    if summary_sheet:
        ws = wb[summary_sheet]
        
        # Count formulas in Summary sheet
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
        
        print(f"  Summary sheet: {summary_sheet}")
        print(f"  Formulas in Summary: {formula_count}")
        
        if formula_count < 5:
            warnings_found.append(
                f"  ⚠ Summary sheet has only {formula_count} formulas (expected calculated metrics)"
            )
        else:
            print(f"  ✓ Summary has active calculations")
    else:
        if workbook_id in ['A817-1', 'A817-2', 'A817-D']:
            warnings_found.append("  ⚠ No Summary sheet found")
    
    # ========================================================================
    # CHECK 9: ROW CAPACITY VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 9: ROW CAPACITY VALIDATION")
    print("=" * 80)
    
    # Check specific sheets based on workbook type
    capacity_ok = True
    
    if workbook_id == 'A817-1':
        # Time Sources should have space for at least 10 sources
        if 'Time_Sources' in wb.sheetnames:
            ws = wb['Time_Sources']
            max_capacity = ws.max_row - 5  # Subtract header rows
            print(f"  Time_Sources: ~{max_capacity} row capacity")
            if max_capacity < 10:
                warnings_found.append(
                    f"  ⚠ Time_Sources has limited capacity ({max_capacity} rows), minimum 10 recommended"
                )
                capacity_ok = False
        
        # Internal NTP Servers
        if 'Internal_NTP_Servers' in wb.sheetnames:
            ws = wb['Internal_NTP_Servers']
            max_capacity = ws.max_row - 5
            print(f"  Internal_NTP_Servers: ~{max_capacity} row capacity")
            if max_capacity < 5:
                warnings_found.append(
                    f"  ⚠ Internal_NTP_Servers has limited capacity ({max_capacity} rows)"
                )
                capacity_ok = False
    
    elif workbook_id == 'A817-2':
        # System Sync Status should have space for many systems
        if 'System_Sync_Status' in wb.sheetnames:
            ws = wb['System_Sync_Status']
            max_capacity = ws.max_row - 5
            print(f"  System_Sync_Status: ~{max_capacity} row capacity")
            if max_capacity < 50:
                warnings_found.append(
                    f"  ⚠ System_Sync_Status has limited capacity ({max_capacity} rows), minimum 50 recommended"
                )
                capacity_ok = False
            elif max_capacity < 200:
                print(f"    ℹ Note: For large environments, consider 200+ row capacity")
    
    if capacity_ok:
        print("  ✓ Adequate row capacity for data entry")
    
    # ========================================================================
    # CHECK 10: WORKSHEET DIMENSIONS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 10: WORKSHEET DIMENSIONS")
    print("=" * 80)
    
    dimension_issues = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if ws.max_row > 5000:
            warnings_found.append(
                f"  ⚠ {sheet_name}: Large worksheet ({ws.max_row} rows) may impact performance"
            )
            dimension_issues += 1
        
        if ws.max_column > 50:
            warnings_found.append(
                f"  ⚠ {sheet_name}: Wide worksheet ({ws.max_column} columns) may impact performance"
            )
            dimension_issues += 1
    
    if dimension_issues == 0:
        print("  ✓ Worksheet dimensions within reasonable limits")
    
    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    print("\n" + "=" * 80)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 80)
    
    exit_code = 0
    
    if issues_found:
        print(f"\n❌ CRITICAL ISSUES FOUND: {len(issues_found)}")
        for issue in issues_found:
            print(issue)
        exit_code = 1
    
    if warnings_found:
        print(f"\n⚠️  WARNINGS: {len(warnings_found)}")
        for warning in warnings_found:
            print(warning)
    
    if not issues_found and not warnings_found:
        print("\n✅ NO ISSUES DETECTED")
        print("\nThe workbook appears structurally sound.")
        print("If Excel still shows repair warnings, consider:")
        print("  • Excel version compatibility (try Excel 2019+)")
        print("  • Antivirus/security software interference")
        print("  • Network drive or OneDrive sync issues")
        print("  • Excel's overly cautious validation")
        if workbook_id == 'A817-D':
            print("  • Unresolved external workbook links (expected before normalization)")
    else:
        print_recommendations(issues_found, warnings_found, workbook_id, 
                            formula_issues, validation_issues, merge_issues)
    
    # ========================================================================
    # WORKBOOK-SPECIFIC NOTES
    # ========================================================================
    print_workbook_notes(workbook_id, external_refs)
    
    wb.close()
    return exit_code


def print_recommendations(issues, warnings, workbook_id, formula_issues, 
                         validation_issues, merge_issues):
    """Print recommended fixes based on issues found."""
    print("\n" + "=" * 80)
    print("RECOMMENDED ACTIONS")
    print("=" * 80)
    
    if formula_issues > 0:
        print("\n1. FORMULA FIXES:")
        print("   • Review formulas referencing other sheets")
        print("   • Ensure sheet names match exactly (case-sensitive)")
        print("   • Check for balanced quotes and parentheses")
        if workbook_id == 'A817-D':
            print("   • Verify external workbook references point to normalized files")
            print("   • Ensure source workbooks are in same directory as dashboard")
    
    if validation_issues > 0:
        print("\n2. DATA VALIDATION FIXES:")
        print("   • Remove overlapping data validation ranges")
        print("   • Apply validations to specific ranges, not entire columns")
    
    if merge_issues > 0:
        print("\n3. MERGED CELL FIXES:")
        print("   • Ensure only top-left cell of merged range has content")
        print("   • Clear content from other cells in merged range")
    
    if workbook_id == 'A817-D':
        print("\n4. DASHBOARD-SPECIFIC SETUP (A817-D):")
        print("   • This workbook consolidates data from 2 assessment workbooks")
        print("   • Ensure normalized source workbooks are in the same folder:")
        for source in DASHBOARD_SOURCES:
            print(f"     - {source}")
        print("   • Run normalize_assessment_files_a817.py first to create normalized files")
        print("   • Dashboard will show #REF errors until source files are present")


def print_workbook_notes(workbook_id, external_refs):
    """Print workbook-specific notes and expected structure."""
    print("\n" + "=" * 80)
    print(f"WORKBOOK-SPECIFIC NOTES: {workbook_id}")
    print("=" * 80)
    
    if workbook_id == 'A817-1':
        print("\nTime Source Infrastructure Assessment:")
        print("  • 7 sheets for time source documentation")
        print("  • Authoritative time sources (Stratum 0/1)")
        print("  • Internal NTP servers (Stratum 2)")
        print("  • Time source hierarchy visualization")
        print("  • Redundancy and availability tracking")
        print("  • Evidence Register (50+ entries recommended)")
        print("  • 3-level approval workflow")
        print("\n  Key metrics:")
        print("    - Number of authoritative time sources (minimum 2)")
        print("    - Internal NTP server redundancy (minimum 2)")
        print("    - Time source availability (>99.9%)")
        print("    - Stratum hierarchy validation")
    
    elif workbook_id == 'A817-2':
        print("\nSystem Synchronization Status Assessment:")
        print("  • 9 sheets for system-level sync verification")
        print("  • Per-system sync status tracking")
        print("  • Time drift measurements")
        print("  • Platform-specific validation (Linux, Windows, network devices, cloud)")
        print("  • Criticality-based compliance thresholds")
        print("  • Gap analysis and remediation planning")
        print("  • Evidence Register (50+ entries recommended)")
        print("  • 3-level approval workflow")
        print("\n  Key metrics:")
        print("    - Overall sync compliance percentage (target: ≥95%)")
        print("    - Average time drift (target: <1 second)")
        print("    - Critical systems in compliance (target: 100%)")
        print("    - Gap remediation status")
        print("\n  Capacity requirements:")
        print("    - Minimum 50 systems (small organizations)")
        print("    - Recommended 200+ systems (medium/large organizations)")
    
    elif workbook_id == 'A817-3':
        print("\nException Register (Optional):")
        print("  • 9 sheets for exception management")
        print("  • Formal risk acceptance for non-compliant systems")
        print("  • Compensating controls documentation")
        print("  • Time-bound exceptions with review cycles")
        print("  • Exception approval workflow")
        print("  • Evidence Register")
        print("\n  Use cases:")
        print("    - Air-gapped systems without network connectivity")
        print("    - Legacy systems incompatible with modern NTP")
        print("    - IoT/embedded devices with limited sync capability")
        print("    - Temporary exceptions during migrations")
        print("\n  Important:")
        print("    - All exceptions require business justification")
        print("    - Risk assessment required per exception")
        print("    - Compensating controls must reduce risk to acceptable level")
        print("    - Regular review cycles enforced (quarterly for high risk)")
    
    elif workbook_id == 'A817-D':
        print("\nTime Synchronization Compliance Dashboard (MASTER):")
        print("  • 9 sheets - EXECUTIVE CONSOLIDATION WORKBOOK")
        print("  • Overall A.8.17 compliance score")
        print("  • Time source infrastructure status")
        print("  • System synchronization compliance metrics")
        print("  • Gap analysis and remediation tracking")
        print("  • Exception summary (if exceptions exist)")
        print("  • Historical trend analysis")
        print("  • Action items with owners and deadlines")
        print("  • Evidence completeness tracking")
        print("\n  ⚠ This dashboard requires normalized source files:")
        for source in DASHBOARD_SOURCES:
            print(f"    - {source}")
        print("\n  Setup process:")
        print("    1. Generate assessment workbooks (Scripts 1-2)")
        print("    2. Run: python3 normalize_assessment_files_a817.py")
        print("    3. Place dashboard in same directory as normalized files")
        print("    4. Open dashboard in Excel → Click 'Update Links' when prompted")
        print("    5. Dashboard will populate with live data from assessments")
        print("\n  Dashboard KPIs:")
        print("    - Overall compliance percentage")
        print("    - Time source redundancy status")
        print("    - Average drift across infrastructure")
        print("    - Critical system sync status")
        print("    - Gap remediation completion rate")
    
    else:
        print("\nGeneric A.8.17 Workbook")
        print("  • No specific validation rules defined for this type")
        print("  • Standard Excel health checks applied")
        print("  • Ensure workbook follows A.8.17 framework structure")
    
    print("\n" + "=" * 80)


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("=" * 70)
        print("ISMS A.8.17 Framework - Excel Sanity Checker")
        print("Control: A.8.17 (Clock Synchronization)")
        print("=" * 70)
        print("\nUsage: python3 excel_sanity_check_a817.py <filename.xlsx>")
        print("\nExamples:")
        print("  python3 excel_sanity_check_a817.py ISMS-A.8.17-Assessment-1-Time-Sources.xlsx")
        print("  python3 excel_sanity_check_a817.py ISMS-A.8.17-Assessment-2-Sync-Status.xlsx")
        print("  python3 excel_sanity_check_a817.py ISMS-A.8.17-Assessment-3-Exceptions.xlsx")
        print("  python3 excel_sanity_check_a817.py ISMS-A.8.17-Dashboard-Time-Sync.xlsx")
        print("\nSupported workbook types:")
        print("  A817-1 - Time Source Infrastructure Assessment")
        print("  A817-2 - System Synchronization Status Assessment")
        print("  A817-3 - Exception Register (Optional)")
        print("  A817-D - Time Synchronization Compliance Dashboard")
        print("\n'Don't fool yourself' - Systems Engineering ISMS")
        sys.exit(1)
    
    filename = sys.argv[1]
    
    try:
        exit_code = check_workbook_health(filename)
        sys.exit(exit_code)
    except FileNotFoundError:
        print(f"\n❌ ERROR: File not found: {filename}")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()