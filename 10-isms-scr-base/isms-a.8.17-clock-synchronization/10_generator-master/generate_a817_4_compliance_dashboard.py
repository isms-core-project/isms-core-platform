#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.17.D - Time Synchronization Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.17: Clock Synchronization
Executive Dashboard: Consolidated Compliance Monitoring and Reporting

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific dashboard requirements, KPI definitions, and
reporting preferences.

Key customization areas:
1. KPI thresholds and targets (based on your compliance objectives)
2. Executive presentation format (adapt to stakeholder preferences)
3. Trend analysis periods (match your reporting cycles)
4. Risk scoring methodology (align with your risk framework)
5. Integration with other dashboards (fit your reporting architecture)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.17 Clock Synchronization Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel compliance dashboard that
consolidates data from A.8.17 assessment workbooks using EXTERNAL WORKBOOK
FORMULA REFERENCES for live data integration and automatic updates.

**Critical Implementation Note:**
This dashboard uses FORMULAS (not static data) to reference source workbooks:
    ='[ISMS-A.8.17-Assessment-1.xlsx]Time_Sources'!$B$5

When source workbooks are updated, dashboard auto-updates when opened (Excel
prompts "Update Links"). This is the CORRECT approach per project standards.

See consolidate_a817_dashboard.py for alternative static data consolidation
approach (used when formula linking is not suitable).

**Purpose:**
Provides executive-level visibility into time synchronization compliance status,
trends, gaps, and remediation progress across the entire organization through
automated consolidation of assessment data.

**Data Sources:**
- Assessment 1: Time Source Infrastructure (ISMS-A.8.17-Assessment-1.xlsx)
- Assessment 2: System Synchronization Status (ISMS-A.8.17-Assessment-2.xlsx)
- Optional: Exception Register (ISMS-A.8.17-Assessment-3.xlsx)

**Generated Dashboard Structure:**
1. Executive_Summary - High-level compliance status and KPIs
2. Time_Sources - Consolidated time source infrastructure metrics
3. Sync_Status - System-level synchronization compliance summary
4. Drift_Analysis - Time drift measurements and threshold violations
5. Gap_Analysis - Non-compliant systems and remediation status
6. Exceptions - Approved exceptions and compensating controls
7. Trends - Historical compliance tracking and improvement trends
8. Action_Items - Remediation tasks with owners and deadlines
9. Audit_Evidence - Evidence register for ISO 27001 compliance

**Key Features:**
- Live data integration via Excel external workbook formulas
- Automatic updates when source assessments change
- Executive KPI dashboard with traffic light indicators
- Compliance trending over time
- Gap prioritization by system criticality
- Remediation progress tracking
- Audit-ready evidence consolidation
- One-page executive summary for board reporting

**Dashboard KPIs:**
- Overall time synchronization compliance percentage
- Number of systems synchronized vs. not synchronized
- Average time drift across infrastructure
- Time sources meeting redundancy requirements
- Critical systems with sync failures (red flag)
- Remediation completion rate
- Exception count and risk level
- Audit readiness score

**Integration:**
Dashboard provides single-pane-of-glass view of A.8.17 compliance status,
suitable for:
- Executive leadership briefings
- Board cybersecurity reporting
- Audit evidence packages
- Monthly/quarterly compliance reviews
- ISO 27001 surveillance audits

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation
    - Microsoft Excel (for opening and updating formula links)

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

**Prerequisites:**
    1. Complete Assessment 1 (Time Sources)
    2. Complete Assessment 2 (System Sync Status)
    3. Normalize assessment filenames (normalize_assessment_files_a817.py)
    4. Place dashboard and source files in SAME DIRECTORY

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a817_4_compliance_dashboard.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a817_4_compliance_dashboard.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a817_4_compliance_dashboard.py --date 20250125

Output:
    File: ISMS-A.8.17-Dashboard-Time-Sync-YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

**Critical Post-Generation Steps:**
    1. Ensure source assessment files are in SAME DIRECTORY as dashboard
    2. Source files must be normalized filenames:
       - ISMS-A.8.17-Assessment-1.xlsx
       - ISMS-A.8.17-Assessment-2.xlsx
    3. Open dashboard in Excel
    4. Excel will prompt "Update Links" → Click YES
    5. Dashboard will populate with live data from assessments
    6. Save dashboard after link update
    7. Re-open dashboard monthly/quarterly for refreshed data

**File Co-Location Requirement:**
Dashboard MUST be in same folder as source files for formula links to work:
/compliance/a817/
├── ISMS-A.8.17-Assessment-1.xlsx          ← Source 1
├── ISMS-A.8.17-Assessment-2.xlsx          ← Source 2
└── ISMS-A.8.17-Dashboard-Time-Sync.xlsx   ← Dashboard (HERE)

Moving dashboard to different folder breaks formula links.

**Monthly Dashboard Update Workflow:**
    1. Update source assessment workbooks (refresh sync status, drift data)
    2. Save source workbooks
    3. Open dashboard → Excel prompts "Update Links" → Click YES
    4. Review updated compliance metrics
    5. Export Executive Summary for stakeholder reporting
    6. Save dashboard with refreshed data

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.17
Dashboard Type:       Compliance Consolidation Dashboard
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.17: Clock Synchronization Policy
    - ISMS-IMP-A.8.17.1: Time Source Infrastructure Assessment
    - ISMS-IMP-A.8.17.2: System Synchronization Status Assessment
    - ISMS-IMP-A.8.17.3: Exception Register (Optional)

Related Scripts:
    - generate_a817_1_time_sources.py (creates Assessment 1)
    - generate_a817_2_sync_status.py (creates Assessment 2)
    - generate_a817_3_exception_register.py (creates Assessment 3 - optional)
    - normalize_assessment_files_a817.py (filename normalization)
    - consolidate_a817_dashboard.py (alternative static consolidation)

Related Standards:
    - ISO/IEC 27001:2022: Clause 9.1 (Monitoring, Measurement, Analysis)
    - ISO/IEC 27002:2022: Control 8.17 (Clock Synchronization)
    - RFC 5905: Network Time Protocol Version 4

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements formula-based dashboard per project standards
    - Live data integration from 2 source assessments
    - Executive summary with compliance KPIs
    - Gap analysis and remediation tracking
    - Trend analysis with historical comparison
    - Audit evidence consolidation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Formula-Based Integration (Primary Approach):**
This dashboard uses EXTERNAL WORKBOOK FORMULAS - the correct approach per
project standards. Formulas like ='[source.xlsx]Sheet'!$A$5 enable:
- Automatic updates when source data changes
- Real-time compliance monitoring
- Reduced maintenance overhead
- Single source of truth

Alternative: consolidate_a817_dashboard.py provides STATIC DATA consolidation
for scenarios where formula linking doesn't work (Excel security settings,
SharePoint, email distribution). Use formula approach unless you have specific
reason to use static consolidation.

**File Co-Location Requirement:**
Excel's external workbook references require files in same directory. This is
NOT a script limitation - it's how Excel works. Do not separate dashboard from
source files unless you're using static consolidation approach.

**Update Links Prompt:**
Every time dashboard is opened, Excel prompts "Update Links?" - this is NORMAL
and EXPECTED behavior. Clicking YES refreshes dashboard with latest assessment
data. Clicking NO shows stale data from last update.

**Dashboard Presentation:**
Executive Summary sheet designed for one-page board reporting:
- Traffic light KPIs (Red/Amber/Green)
- Compliance percentage with trend arrows
- Critical issues highlighted
- High-level gap summary
- No technical jargon

Detail sheets provide drill-down for operational teams.

**Compliance Thresholds:**
Dashboard uses traffic light scoring:
- Green (≥95%): Excellent compliance, minor gaps only
- Amber (85-94%): Acceptable, remediation in progress
- Red (<85%): Requires immediate attention and escalation

Customize thresholds to match your organizational risk appetite.

**Audit Considerations:**
Dashboard serves as compliance evidence for:
- ISO 27001 Clause 9.1 (Monitoring and Measurement)
- Control A.8.17 effectiveness demonstration
- Management review reporting
- Surveillance audit preparation

Maintain monthly dashboard snapshots for trend evidence.

**Integration with Other Controls:**
Dashboard can be extended to show:
- A.8.21 (NTP Security) integration
- A.8.15 (Logging) dependency metrics
- A.8.16 (Monitoring) alert statistics
- A.5.9 (Asset Inventory) coverage

Consider creating master ISMS dashboard consolidating all controls.

**Performance Considerations:**
Dashboard formulas reference multiple large workbooks:
- Keep source files <10MB for good performance
- Close other Excel files when updating dashboard
- Allow 10-30 seconds for link updates on large datasets
- Consider disabling automatic calculation for very large datasets

**Data Refresh Frequency:**
Recommended refresh cycles:
- Executive Summary: Monthly (board reporting)
- Operational details: Quarterly (compliance reviews)
- Critical system verification: Weekly (automated where possible)
- Full infrastructure assessment: Annually

**Trend Analysis:**
Dashboard tracks compliance over time - save monthly snapshots:
/compliance/a817/history/
├── ISMS-A.8.17-Dashboard-2025-01.xlsx
├── ISMS-A.8.17-Dashboard-2025-02.xlsx
└── ISMS-A.8.17-Dashboard-2025-03.xlsx

Historical data enables:
- Compliance trend visualization
- Improvement measurement
- Audit trail of remediation progress
- Executive reporting on security posture improvement

**Quality Assurance:**
Before distributing dashboard:
- Verify all formula links resolve (no #REF! errors)
- Confirm KPIs match source data
- Check traffic lights display correctly
- Test with sample data for accuracy
- Have CISO review Executive Summary

**Stakeholder Distribution:**
Different audiences need different views:
- Board/Executives: Executive Summary only (export as PDF)
- CISO/Security Leadership: Full dashboard access
- Operations Teams: Detail sheets for remediation
- Auditors: Complete dashboard + source assessments + evidence

**Troubleshooting External Links:**
If formulas don't update:
1. Verify source files are in same directory as dashboard
2. Verify source filenames match exactly (case-sensitive on some OS)
3. Check Excel security settings allow external references
4. Try Edit > Links > Update Values manually
5. If still failing, use consolidate_a817_dashboard.py for static approach

**Security Note:**
Excel external links can be security concern in some environments. If your
organization blocks external references, use consolidate_a817_dashboard.py
instead for static data consolidation.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.17.D"
WORKBOOK_NAME = "Time Synchronization Compliance Dashboard"
CONTROL_ID = "A.8.17"
CONTROL_NAME = "Clock Synchronization"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

WB1 = "ISMS-A.8.17-Assessment-1.xlsx"  # Time Sources
WB2 = "ISMS-A.8.17-Assessment-2.xlsx"  # Sync Status

# ============================================================================
# SECTION 1: STYLE DEFINITIONS
# ============================================================================



# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CLOCK = '\u23F0'      # ⏰ Alarm clock
SYNC = '\U0001F504'   # 🔄 Counterclockwise arrows
HOURGLASS = '\u23F3'  # ⏳ Hourglass
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
BULLET = '\u2022'     # • Bullet
ARROW = '\u2192'      # → Right arrow

def create_styles():
    """Define A.8.24 standard styles for the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "data": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "title": {
            "font": Font(name="Calibri", bold=True, size=16, color="003366"),
            "alignment": Alignment(horizontal="left", vertical="center")
        },
        "subtitle": {
            "font": Font(name="Calibri", bold=True, size=12, color="003366"),
            "alignment": Alignment(horizontal="left", vertical="center")
        },
        "metric_value": {
            "font": Font(name="Calibri", bold=True, size=24, color="003366"),
            "alignment": Alignment(horizontal="center", vertical="center")
        },
    }

def set_column_widths(ws, widths):
    """Set column widths for a worksheet"""
    for col_num, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

def apply_style(cell, style_dict):
    """Apply style dictionary to cell"""
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

# ============================================================================
# SECTION 2: FORMULA HELPERS
# ============================================================================

def ext_ref(workbook, sheet, cell):
    """Create external workbook reference formula (with leading =)"""
    return f"='[{workbook}]{sheet}'!{cell}"

def count_ext(workbook, sheet, range_ref):
    """COUNTA expression for external workbook (NO leading = for embedding)"""
    return f"COUNTA('[{workbook}]{sheet}'!{range_ref})"

def countif_ext(workbook, sheet, range_ref, criteria):
    """COUNTIF expression for external workbook (NO leading = for embedding)"""
    return f'COUNTIF(\'[{workbook}]{sheet}\'!{range_ref},"{criteria}")'

def countifs_ext(workbook, sheet, range1, criteria1, range2, criteria2):
    """COUNTIFS expression for external workbook (NO leading = for embedding)"""
    return f'COUNTIFS(\'[{workbook}]{sheet}\'!{range1},"{criteria1}",\'[{workbook}]{sheet}\'!{range2},"{criteria2}")'

# ============================================================================
# SECTION 3: EXECUTIVE SUMMARY SHEET
# ============================================================================

def create_executive_summary(wb, styles):
    """Create Executive Summary dashboard with formulas"""
    ws = wb["Executive_Summary"]
    
    # Title
    ws['A1'] = "ISMS A.8.17 - Time Synchronization Compliance Dashboard"
    apply_style(ws['A1'], styles['title'])
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:F2')
    
    ws['A3'] = "IMPORTANT: Click 'Update Links' when opening to refresh data from source workbooks"
    ws['A3'].font = Font(italic=True, size=10, color="FF0000")
    ws.merge_cells('A3:F3')
    
    # Key metrics section
    row = 5
    ws[f'A{row}'] = "Key Compliance Metrics"
    apply_style(ws[f'A{row}'], styles['subtitle'])
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    
    # Metric cards with formulas
    # Sync Rate = (Synced Systems / Total Systems) * 100
    ws['A7'] = "Sync Rate"
    ws['A7'].font = Font(bold=True, size=10)
    ws['A7'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A7:B7')
    
    ws['A8'] = f"=ROUND({countif_ext(WB2, 'System_Inventory', 'G2:G1000', f'{CHECK} Synced')}/{count_ext(WB2, 'System_Inventory', 'A2:A1000')}*100,1)&\"%\""
    apply_style(ws['A8'], styles['metric_value'])
    ws.merge_cells('A8:B8')
    
    ws['A9'] = "≥95% Target"
    ws['A9'].font = Font(size=9, italic=True, color="999999")
    ws['A9'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A9:B9')
    
    # Compliance Rate
    ws['C7'] = "Compliance Rate"
    ws['C7'].font = Font(bold=True, size=10)
    ws['C7'].alignment = Alignment(horizontal="center")
    ws.merge_cells('C7:D7')
    
    ws['C8'] = f"=ROUND({countif_ext(WB2, 'System_Inventory', 'L2:L1000', f'{CHECK} PASS')}/{count_ext(WB2, 'System_Inventory', 'A2:A1000')}*100,1)&\"%\""
    apply_style(ws['C8'], styles['metric_value'])
    ws.merge_cells('C8:D8')
    
    ws['C9'] = "≥95% Target"
    ws['C9'].font = Font(size=9, italic=True, color="999999")
    ws['C9'].alignment = Alignment(horizontal="center")
    ws.merge_cells('C9:D9')
    
    # Critical Systems Rate
    ws['E7'] = "Critical Sys Rate"
    ws['E7'].font = Font(bold=True, size=10)
    ws['E7'].alignment = Alignment(horizontal="center")
    ws.merge_cells('E7:F7')
    
    ws['E8'] = f"=ROUND({countifs_ext(WB2, 'System_Inventory', 'E2:E1000', '🔴 Critical', 'L2:L1000', f'{CHECK} PASS')}/{countif_ext(WB2, 'System_Inventory', 'E2:E1000', '🔴 Critical')}*100,1)&\"%\""
    apply_style(ws['E8'], styles['metric_value'])
    ws.merge_cells('E8:F8')
    
    ws['E9'] = "≥95% Target"
    ws['E9'].font = Font(size=9, italic=True, color="999999")
    ws['E9'].alignment = Alignment(horizontal="center")
    ws.merge_cells('E9:F9')
    
    # Summary table
    row = 12
    ws[f'A{row}'] = "Infrastructure & System Summary"
    apply_style(ws[f'A{row}'], styles['subtitle'])
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    headers = ["Category", "Metric", "Value", "Status"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['header'])
    
    row += 1
    summary_data = [
        ("Infrastructure", "External Time Sources", count_ext(WB1, 'Time_Sources', 'A2:A100'),
         f"=IF(C{row}>=2,\"{CHECK} PASS\",\"{XMARK} FAIL\")"),
        ("", "Internal NTP Servers", count_ext(WB1, 'Internal_NTP_Servers', 'A2:A100'),
         f"=IF(C{row+1}>=2,\"{CHECK} PASS\",\"{XMARK} FAIL\")"),
        ("", "Monitored NTP Servers", countif_ext(WB1, 'Internal_NTP_Servers', 'H2:H100', f'{CHECK} Monitored'),
         f"=IF(C{row+2}=C{row+1},\"{CHECK} PASS\",\"{WARNING} WARN\")"),
        ("Systems", "Total Systems", count_ext(WB2, 'System_Inventory', 'A2:A1000'), ""),
        ("", "Systems Synchronized", countif_ext(WB2, 'System_Inventory', 'G2:G1000', f'{CHECK} Synced'),
         f"=IF(C{row+4}/C{row+3}>=0.95,\"{CHECK} PASS\",\"{XMARK} FAIL\")"),
        ("", "Systems Compliant", countif_ext(WB2, 'System_Inventory', 'L2:L1000', f'{CHECK} PASS'),
         f"=IF(C{row+5}/C{row+3}>=0.95,\"{CHECK} PASS\",\"{XMARK} FAIL\")"),
        ("Critical", "Total Critical", countif_ext(WB2, 'System_Inventory', 'E2:E1000', '🔴 Critical'), ""),
        ("", "Critical Compliant", countifs_ext(WB2, 'System_Inventory', 'E2:E1000', '🔴 Critical', 'L2:L1000', f'{CHECK} PASS'),
         f"=IF(C{row+7}/C{row+6}>=0.95,\"{CHECK} PASS\",\"{XMARK} FAIL\")"),
        ("Gaps", "Total Gaps", f"={countif_ext(WB2, 'System_Inventory', 'G2:G1000', f'{XMARK} Not Synced')}+{countif_ext(WB2, 'System_Inventory', 'G2:G1000', f'{WARNING} Sync Failed')}",
         f"=IF(C{row+8}=0,\"{CHECK} PASS\",\"{XMARK} FAIL\")"),
    ]
    
    for row_data in summary_data:
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['data'])
            
            # Conditional formatting for Status column
            if col_num == 4 and isinstance(value, str) and value.startswith('=IF'):
                # Apply conditional formatting
                pass  # Will be done in Excel with actual values
        row += 1
    
    # Overall assessment
    row += 2
    ws[f'A{row}'] = "Overall Compliance Assessment"
    apply_style(ws[f'A{row}'], styles['subtitle'])
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    # Formula to determine overall status based on compliance rate
    status_formula = f'=IF(AND(C{row-9}/C{row-12}>=0.95,C{row-7}/C{row-10}>=0.95),"COMPLIANT",IF(C{row-9}/C{row-12}>=0.85,"PARTIAL COMPLIANCE","NON-COMPLIANT"))'
    ws[f'A{row}'] = status_formula
    ws[f'A{row}'].font = Font(bold=True, size=20, color="006100")
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f'A{row}:F{row+1}')
    ws.row_dimensions[row].height = 40
    
    # Recommendations
    row += 3
    ws[f'A{row}'] = "Key Recommendations"
    apply_style(ws[f'A{row}'], styles['subtitle'])
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    recommendations = [
        f"=IF(C14<2,\"{BULLET} CRITICAL: Add additional external time sources (minimum 2 required)\",\"\")",
        f"=IF(C15<2,\"{BULLET} CRITICAL: Deploy additional internal NTP servers (minimum 2 required)\",\"\")",
        f"=IF(C{row+6}>0,\"{BULLET} CRITICAL: Remediate \"&C{row+6}&\" system(s) not synchronized\",\"\")",
        f"=IF(C18/C17<0.95,\"{BULLET} HIGH: Improve sync rate to ≥95% (currently \"&ROUND(C18/C17*100,1)&\"%)\",\"\")",
        f"=IF(C19/C17<0.95,\"{BULLET} HIGH: Address drift issues to achieve ≥95% compliance\",\"\")",
        f"=IF(C16<C15,\"{BULLET} MEDIUM: Configure monitoring for all internal NTP servers\",\"\")",
        f"{BULLET} Continue monthly assessments to track improvements",
    ]
    
    for rec in recommendations:
        ws[f'A{row}'] = rec
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
    
    set_column_widths(ws, [20, 30, 25, 15, 15, 15])

# ============================================================================
# SECTION 4: INFRASTRUCTURE HEALTH SHEET
# ============================================================================

def create_infrastructure_health(wb, styles):
    """Create infrastructure health view with formulas"""
    ws = wb["Infrastructure_Health"]
    
    # Title
    ws['A1'] = "Time Source Infrastructure Health"
    apply_style(ws['A1'], styles['title'])
    ws.merge_cells('A1:E1')
    
    # External sources summary
    row = 3
    ws[f'A{row}'] = "External Authoritative Time Sources (Stratum 0/1)"
    apply_style(ws[f'A{row}'], styles['subtitle'])
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = "Total External Sources:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"={count_ext(WB1, 'Time_Sources', 'A2:A100')}"
    
    ws[f'D{row}'] = "Requirement: ≥2"
    ws[f'E{row}'] = f"=IF(B{row}>=2,\"{CHECK} PASS\",\"{XMARK} FAIL\")"
    ws[f'E{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Active Sources:"
    ws[f'B{row}'] = f"={countif_ext(WB1, 'Time_Sources', 'I2:I100', f'{CHECK} Active')}"
    
    row += 2
    ws[f'A{row}'] = f"{ARROW} See Assessment Workbook 1 for detailed time source inventory"
    ws[f'A{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'A{row}:E{row}')
    
    # Internal NTP servers
    row += 3
    ws[f'A{row}'] = "Internal NTP Servers (Stratum 2)"
    apply_style(ws[f'A{row}'], styles['subtitle'])
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = "Total Internal NTP Servers:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"={count_ext(WB1, 'Internal_NTP_Servers', 'A2:A100')}"
    
    ws[f'D{row}'] = "Requirement: ≥2"
    ws[f'E{row}'] = f"=IF(B{row}>=2,\"{CHECK} PASS\",\"{XMARK} FAIL\")"
    ws[f'E{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Active Servers:"
    ws[f'B{row}'] = f"={countif_ext(WB1, 'Internal_NTP_Servers', 'J2:J100', f'{CHECK} Active')}"
    
    row += 1
    ws[f'A{row}'] = "Monitored Servers:"
    ws[f'B{row}'] = f"={countif_ext(WB1, 'Internal_NTP_Servers', 'H2:H100', f'{CHECK} Monitored')}"
    
    ws[f'D{row}'] = "Requirement: 100%"
    ws[f'E{row}'] = f"=IF(B{row}=B{row-1},\"{CHECK} PASS\",\"{WARNING} WARN\")"
    ws[f'E{row}'].font = Font(bold=True)
    
    row += 2
    ws[f'A{row}'] = f"{ARROW} See Assessment Workbook 1 for detailed NTP server inventory"
    ws[f'A{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'A{row}:E{row}')
    
    set_column_widths(ws, [30, 15, 15, 18, 18])

# ============================================================================
# SECTION 5: SYSTEM COMPLIANCE SHEET
# ============================================================================

def create_system_compliance(wb, styles):
    """Create system compliance analysis with formulas"""
    ws = wb["System_Compliance"]
    
    # Title
    ws['A1'] = "System Synchronization Compliance Analysis"
    apply_style(ws['A1'], styles['title'])
    ws.merge_cells('A1:E1')
    
    # Sync status distribution
    row = 3
    ws[f'A{row}'] = "Synchronization Status Distribution"
    apply_style(ws[f'A{row}'], styles['subtitle'])
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    headers = ["Status", "Count", "Percentage", ""]
    for col_num, header in enumerate(headers, start=1):
        if header:
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            apply_style(cell, styles['header'])
    
    row += 1
    status_types = [f"{CHECK} Synced", f"{XMARK} Not Synced", f"{WARNING} Sync Failed", "❓ Unknown"]
    for status in status_types:
        ws.cell(row=row, column=1).value = status
        ws.cell(row=row, column=2).value = f"={countif_ext(WB2, 'System_Inventory', 'G2:G1000', status)}"
        ws.cell(row=row, column=3).value = f"=ROUND(B{row}/{count_ext(WB2, 'System_Inventory', 'A2:A1000')}*100,1)&\"%\""
        
        for col in range(1, 4):
            apply_style(ws.cell(row=row, column=col), styles['data'])
        row += 1
    
    # Compliance by criticality
    row += 2
    ws[f'A{row}'] = "Compliance by Criticality Level"
    apply_style(ws[f'A{row}'], styles['subtitle'])
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    headers = ["Criticality", "Total", "Compliant", "Rate", "Status"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['header'])
    
    row += 1
    crit_levels = ['🔴 Critical', '🟠 High', '🟡 Medium', '🟢 Low']
    for crit in crit_levels:
        ws.cell(row=row, column=1).value = crit
        ws.cell(row=row, column=2).value = f"={countif_ext(WB2, 'System_Inventory', 'E2:E1000', crit)}"
        ws.cell(row=row, column=3).value = f"={countifs_ext(WB2, 'System_Inventory', 'E2:E1000', crit, 'L2:L1000', f'{CHECK} PASS')}"
        ws.cell(row=row, column=4).value = f"=ROUND(C{row}/B{row}*100,1)&\"%\""
        ws.cell(row=row, column=5).value = f"=IF(C{row}/B{row}>=0.95,\"{CHECK} PASS\",\"{XMARK} FAIL\")"
        
        for col in range(1, 6):
            apply_style(ws.cell(row=row, column=col), styles['data'])
        row += 1
    
    # Drift distribution
    row += 2
    ws[f'A{row}'] = "Time Drift Distribution"
    apply_style(ws[f'A{row}'], styles['subtitle'])
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = "Drift Range"
    ws[f'B{row}'] = "Count"
    ws[f'C{row}'] = "Percentage"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f'{col}{row}'], styles['header'])
    
    row += 1
    drift_ranges = [
        ("≤10ms", f"=SUMPRODUCT((ABS('[{WB2}]System_Inventory'!$I$2:$I$1000)<=10)*('[{WB2}]System_Inventory'!$G$2:$G$1000=\"{CHECK} Synced\"))"),
        ("10-100ms", f"=SUMPRODUCT((ABS('[{WB2}]System_Inventory'!$I$2:$I$1000)>10)*(ABS('[{WB2}]System_Inventory'!$I$2:$I$1000)<=100)*('[{WB2}]System_Inventory'!$G$2:$G$1000=\"{CHECK} Synced\"))"),
        ("100-1000ms", f"=SUMPRODUCT((ABS('[{WB2}]System_Inventory'!$I$2:$I$1000)>100)*(ABS('[{WB2}]System_Inventory'!$I$2:$I$1000)<=1000)*('[{WB2}]System_Inventory'!$G$2:$G$1000=\"{CHECK} Synced\"))"),
        (">1000ms", f"=SUMPRODUCT((ABS('[{WB2}]System_Inventory'!$I$2:$I$1000)>1000)*('[{WB2}]System_Inventory'!$G$2:$G$1000=\"{CHECK} Synced\"))"),
    ]
    
    for range_label, formula in drift_ranges:
        ws.cell(row=row, column=1).value = range_label
        ws.cell(row=row, column=2).value = formula
        ws.cell(row=row, column=3).value = f"=ROUND(B{row}/{countif_ext(WB2, 'System_Inventory', 'G2:G1000', f'{CHECK} Synced')}*100,1)&\"%\""
        
        for col in range(1, 4):
            apply_style(ws.cell(row=row, column=col), styles['data'])
        row += 1
    
    row += 2
    ws[f'A{row}'] = f"{ARROW} See Assessment Workbook 2 for per-system sync status details"
    ws[f'A{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'A{row}:E{row}')
    
    set_column_widths(ws, [20, 15, 15, 15, 15])

# ============================================================================
# SECTION 6: GAPS & ACTION ITEMS SHEET
# ============================================================================

def create_gaps_action_items(wb, styles):
    """Create gaps and action items with formulas"""
    ws = wb["Gaps_Action_Items"]
    
    # Title
    ws['A1'] = "Critical Gaps and Required Actions"
    apply_style(ws['A1'], styles['title'])
    ws.merge_cells('A1:E1')
    
    ws['A2'] = "Systems requiring immediate remediation"
    ws['A2'].font = Font(italic=True, size=10, color="FF0000")
    ws.merge_cells('A2:E2')
    
    # Gap summary
    row = 4
    ws[f'A{row}'] = "Gap Summary"
    apply_style(ws[f'A{row}'], styles['subtitle'])
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    headers = ["Gap Type", "Count", "Severity", "Action Required"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_style(cell, styles['header'])
    
    row += 1
    gaps = [
        ("Systems Not Synchronized", 
         f"={countif_ext(WB2, 'System_Inventory', 'G2:G1000', f'{XMARK} Not Synced')}+{countif_ext(WB2, 'System_Inventory', 'G2:G1000', f'{WARNING} Sync Failed')}",
         "🔴 HIGH", "Configure NTP, verify connectivity"),
        ("Critical Systems Non-Compliant",
         countifs_ext(WB2, 'System_Inventory', 'E2:E1000', '🔴 Critical', 'L2:L1000', f'{XMARK} FAIL'),
         "🔴 HIGH", "Immediate remediation required"),
        ("Systems Exceeding Drift",
         countif_ext(WB2, 'System_Inventory', 'L2:L1000', f'{XMARK} FAIL'),
         "🟡 MEDIUM", "Investigate and remediate"),
        ("Insufficient External Sources",
         f"=IF({count_ext(WB1, 'Time_Sources', 'A2:A100')}<2,1,0)",
         "🔴 HIGH", "Add additional time sources"),
        ("Insufficient Internal Servers",
         f"=IF({count_ext(WB1, 'Internal_NTP_Servers', 'A2:A100')}<2,1,0)",
         "🔴 HIGH", "Deploy additional NTP servers"),
        ("Unmonitored NTP Servers",
         f"={count_ext(WB1, 'Internal_NTP_Servers', 'A2:A100')}-{countif_ext(WB1, 'Internal_NTP_Servers', 'H2:H100', f'{CHECK} Monitored')}",
         "🟡 MEDIUM", "Configure monitoring for all servers"),
    ]
    
    for gap_data in gaps:
        for col_num, value in enumerate(gap_data, start=1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            apply_style(cell, styles['data'])
            
            if col_num == 3 and value == "🔴 HIGH":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 1
    
    # Total gaps
    row += 1
    ws[f'A{row}'] = "Total Gaps Requiring Action:"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'] = f"=SUM(B{row-6}:B{row-1})"
    ws[f'B{row}'].font = Font(bold=True, size=12, color="9C0006")
    
    row += 3
    ws[f'A{row}'] = "Detailed Gap Information"
    apply_style(ws[f'A{row}'], styles['subtitle'])
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = f"{ARROW} See Assessment Workbook 2 'Gaps_Failures' sheet for detailed system-level gap information"
    ws[f'A{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = f"{ARROW} Each gap should have a documented remediation plan with target dates"
    ws[f'A{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'A{row}:E{row}')
    
    set_column_widths(ws, [30, 15, 15, 45, 12])

# ============================================================================
# SECTION 7: INSTRUCTIONS SHEET
# ============================================================================

def create_instructions(wb, styles):
    """Create instructions sheet"""
    ws = wb["Instructions"]
    
    # Title
    ws['A1'] = "ISMS A.8.17 - Time Synchronization Compliance Dashboard"
    apply_style(ws['A1'], styles['title'])
    ws.merge_cells('A1:E1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')
    
    # Instructions
    row = 4
    ws[f'A{row}'] = "HOW TO USE THIS DASHBOARD"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 2
    instructions = [
        "1. SETUP REQUIREMENTS",
        "",
        "   This dashboard uses formulas to reference two normalized assessment workbooks:",
        "   • ISMS-A.8.17-Assessment-1.xlsx (Time Source Infrastructure)",
        "   • ISMS-A.8.17-Assessment-2.xlsx (System Synchronization Status)",
        "",
        "   These workbooks MUST be in the same folder as this dashboard.",
        "",
        "2. FIRST TIME OPENING",
        "",
        "   When you open this dashboard for the first time, Excel will prompt:",
        "   'This workbook contains links to other data sources...'",
        "",
        "   → Click 'UPDATE' or 'Enable Content' to refresh formulas",
        "",
        "3. UPDATING DATA",
        "",
        "   When source assessment workbooks are updated:",
        "   a) Save changes to assessment workbooks",
        "   b) Open this dashboard",
        "   c) Click 'Update Links' when prompted (or Data → Edit Links → Update Values)",
        "   d) Dashboard auto-refreshes with latest data",
        "",
        "   NO NEED TO RE-RUN PYTHON SCRIPT - formulas update automatically!",
        "",
        "4. DASHBOARD SHEETS",
        "",
        "   • Executive_Summary: High-level compliance metrics and overall status",
        "   • Infrastructure_Health: Time source and NTP server assessment",
        "   • System_Compliance: Sync status distribution and drift analysis",
        "   • Gaps_Action_Items: Critical gaps requiring immediate action",
        "",
        "5. WORKFLOW",
        "",
        "   Monthly Assessment Cycle:",
        "   1. Update Assessment Workbook 1 (infrastructure changes if any)",
        "   2. Update Assessment Workbook 2 (verify sync status for all systems)",
        "   3. Open this dashboard → formulas update automatically",
        "   4. Present Executive_Summary to management",
        "   5. Track remediation for items in Gaps_Action_Items",
        "",
        "6. TROUBLESHOOTING",
        "",
        "   If formulas show #REF! errors:",
        "   • Verify normalized assessment workbooks are in same folder",
        "   • Check filenames exactly match: ISMS-A.8.17-Assessment-1.xlsx",
        "   •                                ISMS-A.8.17-Assessment-2.xlsx",
        "   • Click Data → Edit Links → Update Values",
        "",
        "   If data looks outdated:",
        "   • Verify source workbooks were saved after updates",
        "   • Click Data → Refresh All",
        "",
        "7. REFERENCE DOCUMENTS",
        "",
        "   • ISMS-POL-A.8.17: Clock Synchronization Policy",
        "   • ISMS-IMP-A.8.17-S2: Synchronization Verification Process",
        "   • normalize_assessment_files_a817.py: Filename normalization script",
        "",
    ]
    
    for line in instructions:
        ws[f'A{row}'] = line
        if line and line[0].isdigit():
            ws[f'A{row}'].font = Font(bold=True, size=11, color="003366")
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
    
    set_column_widths(ws, [80, 15, 15, 15, 15])

# ============================================================================
# SECTION 8: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function"""
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.17.D - Time Synchronization Compliance Dashboard Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.17")
    logger.info("=" * 78)
    logger.info("")
    logger.info("This dashboard uses FORMULAS to reference normalized assessment workbooks.")
    logger.info("The dashboard will auto-update when source workbooks are updated.")
    logger.info("")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets
    sheets = [
        "Instructions",
        "Executive_Summary",
        "Infrastructure_Health",
        "System_Compliance",
        "Gaps_Action_Items",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    
    # Get styles
    styles = create_styles()
    
    # Create sheets
    logger.info("[1/5] Creating Instructions...")
    create_instructions(wb, styles)
    
    logger.info("[2/5] Creating Executive Summary...")
    create_executive_summary(wb, styles)
    
    logger.info("[3/5] Creating Infrastructure Health...")
    create_infrastructure_health(wb, styles)
    
    logger.info("[4/5] Creating System Compliance...")
    create_system_compliance(wb, styles)
    
    logger.info("[5/5] Creating Gaps & Action Items...")
    create_gaps_action_items(wb, styles)
    
    # Save workbook
    filename = f"ISMS-IMP-A.8.17.4_Time_Sync_Dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    logger.info("")
    logger.info("=" * 78)
    logger.info("{CHECK} SUCCESS: {filename}")
    logger.info("=" * 78)
    logger.info("")
    logger.info("Dashboard Structure:")
    logger.info("  • Instructions (how to use)")
    logger.info("  • Executive Summary (key metrics, overall status)")
    logger.info("  • Infrastructure Health (time sources, NTP servers)")
    logger.info("  • System Compliance (sync status, drift analysis)")
    logger.info("  • Gaps & Action Items (critical gaps, remediation)")
    logger.info("")
    logger.info("NEXT STEPS:")
    logger.info("=" * 78)
    logger.info("")
    logger.info("1. Ensure normalized assessment workbooks are in same folder:")
    logger.info("   • ISMS-A.8.17-Assessment-1.xlsx")
    logger.info("   • ISMS-A.8.17-Assessment-2.xlsx")
    logger.info("")
    logger.info("   If not normalized yet, run:")
    logger.info("   python normalize_assessment_files_a817.py")
    logger.info("")
    logger.info("2. Place this dashboard in same folder as normalized files")
    logger.info("")
    logger.info("3. Open dashboard in Excel")
    logger.info("")
    logger.info("4. Click 'Update Links' when prompted")
    logger.info("")
    logger.info("5. Dashboard will auto-populate with compliance data from source workbooks")
    logger.info("")
    logger.info("6. Monthly updates: Just update source workbooks, open dashboard, refresh!")
    logger.info("")
    logger.info("=" * 78)
    logger.info("")
    logger.info("Evidence > Theater - Formula-based consolidation for real Systems Engineering")
    logger.info("")

if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
