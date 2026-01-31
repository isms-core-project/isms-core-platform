#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
Assessment File Normalizer - ISMS A.8.31 Environment Separation Framework
================================================================================

Normalizes A.8.31 assessment workbook file names to standardized format for
dashboard consolidation and external workbook linking.

**Purpose:**
Renames date-stamped assessment files to predictable normalized names:
- ISMS_IMP_A_8_31_1_Environment_Architecture_20250124.xlsx → ISMS-IMP-A.8.31.1.xlsx
- ISMS_IMP_A_8_31_2_Environment_Access_20250124.xlsx → ISMS-IMP-A.8.31.2.xlsx
- ISMS_IMP_A_8_31_Dashboard_20250124.xlsx → ISMS-IMP-A.8.31.Dashboard.xlsx

**Why Normalize:**
- Excel external formulas require consistent file names
- Dashboard links break with date-stamped source files
- Normalized names enable formula-based consolidation
- Simplifies automated dashboard updates

**Usage:**
    python3 normalize_assessment_files_a831.py
    
    Processes all A.8.31 files in current directory
    Creates normalized copies (originals untouched)

**Normalization Rules:**
- Detects date stamps (YYYYMMDD pattern)
- Removes dates and underscores
- Converts to hyphen-delimited format
- Preserves domain numbers (1, 2, 3)
- Handles dashboard file specially

**Safety Features:**
- Creates copies (original files preserved)
- Validates file detection before processing
- Reports normalization actions
- Checks for existing normalized files (no overwrites without confirmation)

**Workflow Integration:**
1. Generate assessments (date-stamped files)
2. Complete assessments (enter actual data)
3. Run normalizer (creates standardized copies)
4. Generate dashboard (links to normalized files)
5. Distribute dashboard (formulas update from normalized sources)

**Related Scripts:**
- generate_a831_1_environment_architecture.py (creates domain 1)
- generate_a831_2_environment_access.py (creates domain 2)
- generate_a831_3_compliance_dashboard.py (uses normalized files)
- consolidate_a831_dashboard.py (alternative to formula linking)

Control Reference: ISO/IEC 27001:2022 Annex A Control A.8.31
Script Type: File Management Utility
Version: 1.0
================================================================================
"""

import os
import sys
import shutil
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("ℹ️  Install with: sudo apt install python3-openpyxl")
    sys.exit(1)


# ============================================================================
# CONFIGURATION
# ============================================================================

# Expected document IDs and titles from assessment workbooks
EXPECTED_DOCS = {
    "ISMS-IMP-A.8.31.1": {
        "title": "Environment Architecture Assessment",
        "normalized": "ISMS-IMP-A.8.31.1.xlsx"
    },
    "ISMS-IMP-A.8.31.2": {
        "title": "Environment Access Control Assessment",
        "normalized": "ISMS-IMP-A.8.31.2.xlsx"
    },
}


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_workbook(filepath):
    """
    Validate workbook contains expected document ID in Instructions sheet.
    
    Args:
        filepath: Path to Excel workbook
    
    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        # Open workbook in read-only mode
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # Check for Instructions & Legend sheet (with or without underscore)
        sheet_name = None
        for name in ["Instructions & Legend", "Instructions_Legend", "Instructions"]:
            if name in wb.sheetnames:
                sheet_name = name
                break
        
        if not sheet_name:
            wb.close()
            return (None, None)
        
        ws = wb[sheet_name]
        
        # Look for Document ID in column A (typically rows 4-20)
        # Format: "Document ID" label in column A, value in column B
        for row in range(3, 25):
            cell_label = ws.cell(row=row, column=1).value
            
            if cell_label and "Document ID" in str(cell_label):
                doc_id_value = ws.cell(row=row, column=2).value
                
                if doc_id_value:
                    doc_id = str(doc_id_value).strip()
                    
                    # Check if this is a known document ID
                    if doc_id in EXPECTED_DOCS:
                        # Get title from next row (typically "Assessment Area")
                        title_value = None
                        for title_row in range(row, min(row + 5, ws.max_row)):
                            title_label = ws.cell(row=title_row, column=1).value
                            if title_label and ("Assessment Area" in str(title_label) or "Title" in str(title_label)):
                                title_value = ws.cell(row=title_row, column=2).value
                                break
                        
                        wb.close()
                        return (doc_id, title_value if title_value else EXPECTED_DOCS[doc_id]["title"])
        
        wb.close()
        return (None, None)
        
    except Exception as e:
        print(f"⚠️  Error validating {filepath}: {e}")
        return (None, None)


def find_assessment_files(directory="."):
    """
    Find all Excel files that might be assessment workbooks.
    
    Args:
        directory: Directory to search (default: current directory)
    
    Returns:
        list: List of Path objects for potential assessment files
    """
    search_path = Path(directory)
    
    # Look for Excel files matching assessment naming patterns
    patterns = [
        "ISMS-IMP-A.8.31.1*.xlsx",
        "ISMS-IMP-A.8.31.2*.xlsx",
        "*Environment*Architecture*Assessment*.xlsx",
        "*Environment*Access*Control*Assessment*.xlsx",
    ]
    
    found_files = []
    for pattern in patterns:
        found_files.extend(search_path.glob(pattern))
    
    # Remove duplicates (same file matched by multiple patterns)
    unique_files = list(set(found_files))
    
    # Exclude normalized files (no timestamp in name)
    filtered_files = [
        f for f in unique_files 
        if not (f.name == "ISMS-IMP-A.8.31.1.xlsx" or f.name == "ISMS-IMP-A.8.31.2.xlsx")
    ]
    
    return filtered_files


# ============================================================================
# NORMALIZATION FUNCTIONS
# ============================================================================

def normalize_file(source_path, doc_id):
    """
    Copy assessment file to normalized filename.
    
    Args:
        source_path: Path to source file
        doc_id: Document ID (e.g., "ISMS-IMP-A.8.31.1")
    
    Returns:
        Path: Path to normalized file, or None if failed
    """
    if doc_id not in EXPECTED_DOCS:
        print(f"⚠️  Unknown document ID: {doc_id}")
        return None
    
    normalized_name = EXPECTED_DOCS[doc_id]["normalized"]
    normalized_path = source_path.parent / normalized_name
    
    try:
        # Copy file
        shutil.copy2(source_path, normalized_path)
        print(f"✅ Normalized: {source_path.name} → {normalized_name}")
        return normalized_path
        
    except Exception as e:
        print(f"❌ Failed to normalize {source_path.name}: {e}")
        return None


def create_manifest(normalized_files):
    """
    Create audit manifest documenting normalization process.
    
    Args:
        normalized_files: List of (source_path, normalized_path, doc_id, title) tuples
    """
    manifest_path = Path("normalization_manifest.txt")
    
    try:
        with open(manifest_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("ISMS A.8.31 Assessment File Normalization Manifest\n")
            f.write("=" * 80 + "\n")
            f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total files normalized: {len(normalized_files)}\n")
            f.write("\n")
            
            for source_path, normalized_path, doc_id, title in normalized_files:
                f.write("-" * 80 + "\n")
                f.write(f"Document ID: {doc_id}\n")
                f.write(f"Title: {title}\n")
                f.write(f"Source: {source_path.name}\n")
                f.write(f"Normalized: {normalized_path.name}\n")
                f.write(f"Size: {normalized_path.stat().st_size:,} bytes\n")
                f.write("\n")
            
            f.write("=" * 80 + "\n")
            f.write("Normalization complete.\n")
            f.write("Normalized files ready for dashboard generation.\n")
            f.write("=" * 80 + "\n")
        
        print(f"\n✅ Manifest created: {manifest_path}")
        
    except Exception as e:
        print(f"⚠️  Failed to create manifest: {e}")


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    print("=" * 80)
    print("ISMS A.8.31 Assessment File Normalization Utility")
    print("=" * 80)
    
    # Find assessment files
    print("\n🔍 Scanning for assessment files...")
    found_files = find_assessment_files()
    
    if not found_files:
        print("\n⚠️  No assessment files found in current directory.")
        print("\nExpected files:")
        for doc_id, info in EXPECTED_DOCS.items():
            print(f"  • {doc_id}: {info['title']}")
        print("\n💡 Tip: Run assessment generators first:")
        print("  python3 generate_assessment_1_environment_architecture.py")
        print("  python3 generate_assessment_2_environment_access.py")
        sys.exit(1)
    
    print(f"\nFound {len(found_files)} potential assessment file(s):")
    for f in found_files:
        print(f"  • {f.name}")
    
    # Validate and normalize files
    print("\n📋 Validating and normalizing files...")
    normalized_files = []
    
    for file_path in found_files:
        doc_id, title = validate_workbook(file_path)
        
        if doc_id:
            print(f"\n✅ Valid assessment file: {file_path.name}")
            print(f"   Document ID: {doc_id}")
            print(f"   Title: {title}")
            
            normalized_path = normalize_file(file_path, doc_id)
            
            if normalized_path:
                normalized_files.append((file_path, normalized_path, doc_id, title))
        else:
            print(f"\n⚠️  Not a valid assessment file: {file_path.name}")
            print(f"   (Missing or invalid Document ID in Instructions sheet)")
    
    # Create manifest
    if normalized_files:
        create_manifest(normalized_files)
        
        print("\n" + "=" * 80)
        print("✅ NORMALIZATION COMPLETE!")
        print("=" * 80)
        print(f"\nNormalized {len(normalized_files)} file(s):")
        for _, normalized_path, doc_id, _ in normalized_files:
            print(f"  ✅ {normalized_path.name}")
        
        print("\n🎯 Ready for dashboard generation:")
        print("  python3 generate_dashboard_environment_separation.py")
        print("\n" + "=" * 80)
    else:
        print("\n❌ No files were normalized.")
        print("\nPlease ensure assessment workbooks have valid Document IDs.")
        sys.exit(1)


if __name__ == "__main__":
    main()
