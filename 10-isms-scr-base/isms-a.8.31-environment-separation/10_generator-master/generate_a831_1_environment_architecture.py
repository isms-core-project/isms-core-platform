#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.31.1 - Environment Architecture Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.31: Separation of Development, Test and Production
Assessment Domain 1 of 3: Environment Architecture

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific environment architecture, separation mechanisms,
and infrastructure requirements.

Key customization areas:
1. Environment definitions (dev, test, staging, prod per your landscape)
2. Separation mechanisms (network, infrastructure, cloud accounts)
3. Cloud provider specifics (AWS, Azure, GCP per your deployment)
4. Compliance thresholds (aligned with your risk appetite)
5. Evidence requirements (per your audit framework)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for environment separation)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

Generates Excel assessment workbook for evaluating environment separation
architecture compliance per ISO/IEC 27001:2022 Control A.8.31.

**Purpose:**
Creates structured assessment for environment architecture including:
- Environment definitions (dev, test, staging, production)
- Separation mechanisms (network, infrastructure, data, credentials)
- Network isolation and segmentation verification
- Infrastructure separation compliance (cloud accounts, clusters)

**Assessment Domains:**
Domain 1: Environment Architecture (network, infrastructure separation)
Domain 2: Environment Access Control (who accesses what, production restrictions)
Domain 3: Data Handling & Promotion (no prod data in dev/test, promotion workflow)

**Usage:**
    python3 generate_a831_1_environment_architecture.py
    
    Output: ISMS_IMP_A_8_31_1_Environment_Architecture_YYYYMMDD.xlsx

**Workbook Structure:**
- Executive Summary (compliance overview, critical findings)
- Gap Analysis (separation deficiencies, risk ratings)
- Environment Inventory (dev, test, staging, production definitions)
- Network Separation (VLANs, subnets, firewall rules)
- Infrastructure Separation (cloud accounts, K8s namespaces, VM isolation)
- Data Separation (database segregation, no prod data in dev/test)
- Credential Separation (environment-specific credentials, PAM usage)
- Evidence Register (architecture diagrams, config exports)
- Action Items (remediation tasks, owners, deadlines)

**Integration:**
- Consolidates into A.8.31 Compliance Dashboard
- Links to A.8.25-26-29 (Secure Development), A.8.32 (Change Management)
- Supports A.5.15-16-18 (IAM) environment access control

Control Reference: ISO/IEC 27001:2022 Annex A Control A.8.31
Script Type: Assessment Workbook Generator
Version: 1.0
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.31.1"
WORKBOOK_NAME = "Environment Architecture Assessment"
CONTROL_ID = "A.8.31"
CONTROL_NAME = "Separation of Development, Test and Production Environments"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure for A.8.31.1
    sheets = [
        "Instructions_Legend",
        "Environment_Inventory",
        "Network_Separation",
        "Infrastructure_Separation",
        "Data_Separation",
        "Credential_Separation",
        "Configuration_Consistency",
        "Gap_Analysis",
        "Evidence_Register",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "formula_cell": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
        "status_green": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_yellow": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_red": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "gap_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], "color") else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, "rgb") else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, "rgb") else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS
# ============================================================================

def create_base_validations(ws):
    """Create data validation objects for standard dropdowns."""
    validations = {
        "yes_no": DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        "yes_no_partial": DataValidation(
            type="list",
            formula1='"✅ Yes,❌ No,⚠️ Partial,❓ Unknown"',
            allow_blank=False
        ),
        "compliance_status": DataValidation(
            type="list",
            formula1='"✅ Compliant,❌ Non-Compliant,⚠️ Partial,📋 Not Assessed,➖ N/A"',
            allow_blank=False
        ),
        "environment_type": DataValidation(
            type="list",
            formula1='"Development,Testing/QA,Staging,Production,Sandbox,DR/Backup,Training,Other"',
            allow_blank=False
        ),
        "infrastructure_type": DataValidation(
            type="list",
            formula1='"On-Premises,AWS,Azure,GCP,Hybrid,Multi-Cloud,Kubernetes,Other"',
            allow_blank=False
        ),
        "separation_level": DataValidation(
            type="list",
            formula1='"🟢 Complete,🟡 Partial,🔴 None,➖ N/A"',
            allow_blank=False
        ),
        "risk_severity": DataValidation(
            type="list",
            formula1='"🔴 Critical,🟠 High,🟡 Medium,🟢 Low,⚫ Informational"',
            allow_blank=False
        ),
    }
    
    for val in validations.values():
        ws.add_data_validation(val)
    
    return validations


# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(wb, styles):
    """Create Instructions & Legend sheet."""
    ws = wb["Instructions_Legend"]
    
    # Header
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "ISMS-IMP-A.8.31.1 – Environment Architecture Assessment"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 40
    
    # Subtitle
    ws.merge_cells("A2:F2")
    cell = ws["A2"]
    cell.value = "ISO/IEC 27001:2022 - Control A.8.31: Separation of Development, Test and Production Environments"
    apply_style(cell, styles["subheader"])
    ws.row_dimensions[2].height = 30
    
    # Document Information
    row = 4
    info_items = [
        ("Document ID:", "ISMS-IMP-A.8.31.1"),
        ("Assessment Area:", "Environment Architecture & Separation"),
        ("Related Policy:", "ISMS-POL-A.8.31-S2.1"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Quarterly"),
    ]
    
    for label, value in info_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "USER INPUT" in value:
            apply_style(cell, styles["input_cell"])
        row += 1
    
    # How to Use This Workbook
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws.cell(row=row, column=1, value="How to Use This Workbook")
    apply_style(cell, styles["subheader"])
    
    row += 1
    instructions = [
        "1. Complete the Environment_Inventory sheet with all your environments (dev, test, staging, production)",
        "2. Assess Network_Separation - verify VLANs, VPCs, firewall rules prevent cross-environment access",
        "3. Assess Infrastructure_Separation - verify separate servers, cloud accounts, databases",
        "4. Assess Data_Separation - confirm NO production data in dev/test environments",
        "5. Assess Credential_Separation - verify unique credentials per environment",
        "6. Review Configuration_Consistency - check for drift between staging and production",
        "7. Complete Gap_Analysis - document non-compliance areas and remediation plans",
        "8. Maintain Evidence_Register for audit traceability",
        "9. Obtain final approval and sign-off",
    ]
    
    for instruction in instructions:
        ws.cell(row=row, column=1, value=instruction)
        row += 1
    
    # Status Legend
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws.cell(row=row, column=1, value="Status Legend")
    apply_style(cell, styles["subheader"])
    
    row += 1
    legend = [
        ("✅", "Compliant / Yes / Implemented", "Green (C6EFCE)"),
        ("❌", "Non-Compliant / No / Not Implemented", "Red (FFC7CE)"),
        ("⚠️", "Partial / Warning / Needs Attention", "Yellow (FFEB9C)"),
        ("📋", "Not Assessed / Under Review", "Blue (B4C7E7)"),
        ("❓", "Unknown / To Be Determined", "Gray"),
        ("➖", "N/A / Not Applicable", "Gray"),
    ]
    
    ws.cell(row=row, column=1, value="Symbol").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Status").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Color Code").font = Font(bold=True)
    row += 1
    
    for symbol, status, color in legend:
        ws.cell(row=row, column=1, value=symbol)
        ws.cell(row=row, column=2, value=status)
        ws.cell(row=row, column=3, value=color)
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20


# ============================================================================
# SECTION 4: ENVIRONMENT INVENTORY SHEET
# ============================================================================

def create_environment_inventory_sheet(wb, styles):
    """Create Environment Inventory sheet."""
    ws = wb["Environment_Inventory"]
    
    # Header
    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = "ENVIRONMENT INVENTORY"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:I2")
    cell = ws["A2"]
    cell.value = "Document all environments in use by [Organization]"
    apply_style(cell, styles["subheader"])
    
    # Column headers
    row = 4
    headers = [
        ("A", "Environment Name", 25),
        ("B", "Environment Type", 20),
        ("C", "Purpose", 40),
        ("D", "Infrastructure Type", 20),
        ("E", "Primary Users", 25),
        ("F", "Data Type Allowed", 30),
        ("G", "Availability Target", 20),
        ("H", "Change Control Level", 25),
        ("I", "Notes", 40),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # Data validations
    validations = create_base_validations(ws)
    
    # Apply validations to data rows (5-50)
    validations["environment_type"].add("B5:B50")
    validations["infrastructure_type"].add("D5:D50")
    
    # Sample data row
    row = 5
    sample_data = [
        ("dev-myapp", "Development", "Application development", "AWS", "Developers, DevOps", "Synthetic/Anonymized ONLY", "Best effort", "Minimal", "Dev account: 111111111111"),
        ("test-myapp", "Testing/QA", "QA and UAT", "AWS", "QA team, Business users", "Synthetic/Anonymized ONLY", "95% business hours", "Moderate", "Test account: 222222222222"),
        ("prod-myapp", "Production", "Live customer operations", "AWS", "Operations team ONLY", "Production data", "99.9% SLA", "Maximum (CAB)", "Prod account: 444444444444"),
    ]
    
    for data in sample_data:
        for idx, value in enumerate(data):
            cell = ws.cell(row=row, column=idx+1, value=value)
            apply_style(cell, styles["input_cell"])
        row += 1


# ============================================================================
# SECTION 5: NETWORK SEPARATION SHEET
# ============================================================================

def create_network_separation_sheet(wb, styles):
    """Create Network Separation assessment sheet."""
    ws = wb["Network_Separation"]
    
    # Header
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "NETWORK SEPARATION ASSESSMENT"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = "Verify network-level isolation between environments"
    apply_style(cell, styles["subheader"])
    
    # Column headers
    row = 4
    headers = [
        ("A", "Environment Pair", 30),
        ("B", "Network Segment (VLAN/VPC)", 35),
        ("C", "Firewall Rules", 30),
        ("D", "Cross-Environment Traffic", 30),
        ("E", "Separation Status", 20),
        ("F", "Test Result", 30),
        ("G", "Compliance", 20),
        ("H", "Notes/Evidence", 40),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # Data validations
    validations = create_base_validations(ws)
    validations["separation_level"].add("E5:E50")
    validations["compliance_status"].add("G5:G50")
    
    # Sample data
    row = 5
    sample_data = [
        ("Dev → Test", "VLAN 100 → VLAN 200", "Default DENY", "❌ No traffic allowed", "🟢 Complete", "Ping test FAILED (expected)", "✅ Compliant", "Firewall rule export: evidence_001.pdf"),
        ("Dev → Prod", "VLAN 100 → VLAN 400", "Default DENY", "❌ No traffic allowed", "🟢 Complete", "SSH test FAILED (expected)", "✅ Compliant", "Penetration test: pentest_2026.pdf"),
        ("Test → Prod", "VLAN 200 → VLAN 400", "Default DENY", "❌ No traffic allowed", "🟢 Complete", "HTTP test FAILED (expected)", "✅ Compliant", "Network diagram: network_arch.pdf"),
    ]
    
    for data in sample_data:
        for idx, value in enumerate(data):
            cell = ws.cell(row=row, column=idx+1, value=value)
            apply_style(cell, styles["input_cell"])
        row += 1


# ============================================================================
# SECTION 6: INFRASTRUCTURE SEPARATION SHEET
# ============================================================================

def create_infrastructure_separation_sheet(wb, styles):
    """Create Infrastructure Separation assessment sheet."""
    ws = wb["Infrastructure_Separation"]
    
    # Header
    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = "INFRASTRUCTURE SEPARATION ASSESSMENT"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:I2")
    cell = ws["A2"]
    cell.value = "Verify separate compute, storage, and database resources per environment"
    apply_style(cell, styles["subheader"])
    
    # Column headers
    row = 4
    headers = [
        ("A", "Environment", 20),
        ("B", "Compute Resources", 35),
        ("C", "Database Instances", 35),
        ("D", "Storage/Buckets", 35),
        ("E", "Cloud Account/Sub", 30),
        ("F", "Shared Resources?", 20),
        ("G", "Separation Status", 20),
        ("H", "Compliance", 20),
        ("I", "Evidence", 40),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # Data validations
    validations = create_base_validations(ws)
    validations["yes_no"].add("F5:F50")
    validations["separation_level"].add("G5:G50")
    validations["compliance_status"].add("H5:H50")
    
    # Sample data
    row = 5
    sample_data = [
        ("Development", "3x t3.micro EC2", "dev-db.rds (t3.micro)", "s3://myorg-dev-data", "AWS: 111111111111", "No", "🟢 Complete", "✅ Compliant", "AWS inventory: aws_dev_inventory.xlsx"),
        ("Testing", "3x t3.small EC2", "test-db.rds (t3.small)", "s3://myorg-test-data", "AWS: 222222222222", "No", "🟢 Complete", "✅ Compliant", "AWS inventory: aws_test_inventory.xlsx"),
        ("Production", "10x t3.xlarge EC2", "prod-db.rds (r5.2xlarge)", "s3://myorg-prod-data", "AWS: 444444444444", "No", "🟢 Complete", "✅ Compliant", "AWS inventory: aws_prod_inventory.xlsx"),
    ]
    
    for data in sample_data:
        for idx, value in enumerate(data):
            cell = ws.cell(row=row, column=idx+1, value=value)
            apply_style(cell, styles["input_cell"])
        row += 1


# ============================================================================
# SECTION 7: DATA SEPARATION SHEET
# ============================================================================

def create_data_separation_sheet(wb, styles):
    """Create Data Separation assessment sheet."""
    ws = wb["Data_Separation"]
    
    # Header
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "DATA SEPARATION ASSESSMENT"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = "⚠️ CRITICAL: Verify NO production data in development or testing environments"
    apply_style(cell, styles["subheader"])
    
    # Column headers
    row = 4
    headers = [
        ("A", "Environment", 20),
        ("B", "Data Type Used", 30),
        ("C", "Production Data Present?", 25),
        ("D", "Anonymization Applied?", 25),
        ("E", "Synthetic Data Used?", 25),
        ("F", "Data Source", 35),
        ("G", "Compliance", 20),
        ("H", "Evidence/Notes", 40),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # Data validations
    validations = create_base_validations(ws)
    validations["yes_no"].add("C5:C50")
    validations["yes_no"].add("D5:D50")
    validations["yes_no"].add("E5:E50")
    validations["compliance_status"].add("G5:G50")
    
    # Sample data with CRITICAL check
    row = 5
    sample_data = [
        ("Development", "Synthetic test data", "No", "N/A", "Yes", "Faker library (Python)", "✅ Compliant", "Data generation script: generate_testdata.py"),
        ("Testing", "Anonymized data subset", "No", "Yes", "Partial", "Production DB → anonymization script", "✅ Compliant", "Anonymization procedure: data_anon_v2.py"),
        ("Staging", "Synthetic data", "No", "N/A", "Yes", "Same as dev environment", "✅ Compliant", "Synthetic data generator deployed"),
        ("Production", "Real customer data", "Yes (authorized)", "N/A", "No", "Application database", "✅ Compliant", "Production environment only"),
    ]
    
    for data in sample_data:
        for idx, value in enumerate(data):
            cell = ws.cell(row=row, column=idx+1, value=value)
            apply_style(cell, styles["input_cell"])
            
            # Highlight if production data in non-prod (CRITICAL VIOLATION)
            if idx == 2 and value == "Yes" and data[0] != "Production":
                apply_style(cell, styles["gap_critical"])
        row += 1


# ============================================================================
# SECTION 8: CREDENTIAL SEPARATION SHEET
# ============================================================================

def create_credential_separation_sheet(wb, styles):
    """Create Credential Separation assessment sheet."""
    ws = wb["Credential_Separation"]
    
    # Header
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "CREDENTIAL & SECRETS SEPARATION ASSESSMENT"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = "Verify unique credentials per environment and production secrets in PAM vault"
    apply_style(cell, styles["subheader"])
    
    # Column headers
    row = 4
    headers = [
        ("A", "Credential Type", 30),
        ("B", "Development", 30),
        ("C", "Testing", 30),
        ("D", "Production", 30),
        ("E", "Shared Credentials?", 20),
        ("F", "Prod in PAM Vault?", 20),
        ("G", "Compliance", 20),
        ("H", "Evidence", 40),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # Data validations
    validations = create_base_validations(ws)
    validations["yes_no"].add("E5:E50")
    validations["yes_no"].add("F5:F50")
    validations["compliance_status"].add("G5:G50")
    
    # Sample data
    row = 5
    sample_data = [
        ("Database Admin Password", "dev_admin / [unique-1]", "test_admin / [unique-2]", "prod_admin / [PAM-vault]", "No", "Yes", "✅ Compliant", "CyberArk vault screenshot"),
        ("Application DB User", "dev_app / [unique-3]", "test_app / [unique-4]", "prod_app / [PAM-vault]", "No", "Yes", "✅ Compliant", "Password rotation log"),
        ("AWS Access Key", "AKIA...DEV", "AKIA...TEST", "AKIA...PROD", "No", "Yes", "✅ Compliant", "IAM user list per account"),
        ("SSH Private Key", "/keys/dev_rsa", "/keys/test_rsa", "[PAM vault]", "No", "Yes", "✅ Compliant", "Key management procedure"),
        ("API Key (3rd party)", "sk_test_abc123", "sk_test_xyz789", "sk_live_prod456", "No", "Yes", "✅ Compliant", "Secrets Manager: api-keys"),
    ]
    
    for data in sample_data:
        for idx, value in enumerate(data):
            cell = ws.cell(row=row, column=idx+1, value=value)
            apply_style(cell, styles["input_cell"])
            
            # Highlight shared credentials (VIOLATION)
            if idx == 4 and value == "Yes":
                apply_style(cell, styles["status_red"])
        row += 1


# ============================================================================
# SECTION 9: CONFIGURATION CONSISTENCY SHEET
# ============================================================================

def create_configuration_consistency_sheet(wb, styles):
    """Create Configuration Consistency assessment sheet."""
    ws = wb["Configuration_Consistency"]
    
    # Header
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "CONFIGURATION CONSISTENCY CHECK"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:G2")
    cell = ws["A2"]
    cell.value = "Verify staging mirrors production configuration (drift detection)"
    apply_style(cell, styles["subheader"])
    
    # Column headers
    row = 4
    headers = [
        ("A", "Configuration Item", 35),
        ("B", "Staging Config", 35),
        ("C", "Production Config", 35),
        ("D", "Match?", 15),
        ("E", "Drift %", 15),
        ("F", "Compliance", 20),
        ("G", "Notes", 40),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # Data validations
    validations = create_base_validations(ws)
    validations["yes_no"].add("D5:D50")
    validations["compliance_status"].add("F5:F50")
    
    # Sample data
    row = 5
    sample_data = [
        ("Application Version", "v2.3.1", "v2.3.1", "Yes", "0%", "✅ Compliant", "Versions match"),
        ("Instance Type", "t3.xlarge", "t3.xlarge", "Yes", "0%", "✅ Compliant", "Same compute size"),
        ("Security Group Rules", "5 rules", "5 rules", "Yes", "0%", "✅ Compliant", "Rules identical"),
        ("Environment Variables", "12 vars", "12 vars", "Yes", "0%", "✅ Compliant", "Env vars match"),
        ("Database Schema", "v2.3.1", "v2.3.1", "Yes", "0%", "✅ Compliant", "Schema migration applied"),
    ]
    
    for data in sample_data:
        for idx, value in enumerate(data):
            cell = ws.cell(row=row, column=idx+1, value=value)
            apply_style(cell, styles["input_cell"])
        row += 1


# ============================================================================
# SECTION 10: GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis_sheet(wb, styles):
    """Create Gap Analysis sheet."""
    ws = wb["Gap_Analysis"]
    
    # Header
    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = "GAP ANALYSIS & REMEDIATION PLAN"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:I2")
    cell = ws["A2"]
    cell.value = "Document non-compliance areas and remediation actions"
    apply_style(cell, styles["subheader"])
    
    # Column headers
    row = 4
    headers = [
        ("A", "Gap ID", 12),
        ("B", "Area", 25),
        ("C", "Gap Description", 40),
        ("D", "Risk Severity", 20),
        ("E", "Current State", 30),
        ("F", "Target State", 30),
        ("G", "Remediation Action", 40),
        ("H", "Owner", 20),
        ("I", "Target Date", 15),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # Data validations
    validations = create_base_validations(ws)
    validations["risk_severity"].add("D5:D50")
    
    # Sample gap (if needed)
    row = 5
    ws.cell(row=row, column=1, value="GAP-001")
    ws.cell(row=row, column=2, value="Network Separation")
    ws.cell(row=row, column=3, value="VPC peering exists between dev and staging")
    ws.cell(row=row, column=4, value="🟡 Medium")
    ws.cell(row=row, column=5, value="Dev VPC peered to Staging VPC")
    ws.cell(row=row, column=6, value="No VPC peering between non-prod envs")
    ws.cell(row=row, column=7, value="Remove VPC peering connection vpc-peer-12345")
    ws.cell(row=row, column=8, value="Cloud Architect")
    ws.cell(row=row, column=9, value="2026-02-15")
    
    for col in range(1, 10):
        apply_style(ws.cell(row=row, column=col), styles["input_cell"])


# ============================================================================
# SECTION 11: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register_sheet(wb, styles):
    """Create Evidence Register sheet."""
    ws = wb["Evidence_Register"]
    
    # Header
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "EVIDENCE REGISTER"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:G2")
    cell = ws["A2"]
    cell.value = "Maintain audit trail of all supporting evidence"
    apply_style(cell, styles["subheader"])
    
    # Column headers
    row = 4
    headers = [
        ("A", "Evidence ID", 15),
        ("B", "Evidence Type", 30),
        ("C", "Description", 40),
        ("D", "Related Requirement", 30),
        ("E", "File Location", 40),
        ("F", "Collected Date", 15),
        ("G", "Collected By", 20),
    ]
    
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width
    
    # Sample evidence
    row = 5
    evidence = [
        ("EV-001", "Network Diagram", "Architecture diagram showing VLANs per environment", "Network Separation", "/evidence/network_diagram_2026.pdf", "2026-01-10", "Network Admin"),
        ("EV-002", "Firewall Rules Export", "Firewall configuration showing default deny", "Network Separation", "/evidence/firewall_rules.txt", "2026-01-10", "Security Engineer"),
        ("EV-003", "AWS Account List", "List of AWS accounts per environment", "Infrastructure Separation", "/evidence/aws_accounts.xlsx", "2026-01-11", "Cloud Architect"),
        ("EV-004", "Penetration Test Report", "Network isolation validation", "Network Separation", "/evidence/pentest_2026_Q1.pdf", "2026-01-05", "External Auditor"),
        ("EV-005", "PAM Vault Screenshot", "Production credentials stored in CyberArk", "Credential Separation", "/evidence/pam_vault_prod.png", "2026-01-11", "IAM Admin"),
    ]
    
    for data in evidence:
        for idx, value in enumerate(data):
            cell = ws.cell(row=row, column=idx+1, value=value)
            apply_style(cell, styles["input_cell"])
        row += 1


# ============================================================================
# SECTION 12: APPROVAL SIGN-OFF SHEET
# ============================================================================

def create_approval_signoff_sheet(wb, styles):
    """Create Approval Sign-Off sheet."""
    ws = wb["Approval_Sign_Off"]
    
    # Header
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = "APPROVAL & SIGN-OFF"
    apply_style(cell, styles["header"])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:E2")
    cell = ws["A2"]
    cell.value = "Formal approval of environment architecture assessment"
    apply_style(cell, styles["subheader"])
    
    # Assessment Summary
    row = 4
    ws.cell(row=row, column=1, value="Assessment Summary:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Total Environments Assessed:")
    ws.cell(row=row, column=2, value="[COUNT]")
    row += 1
    ws.cell(row=row, column=1, value="Compliant Environments:")
    ws.cell(row=row, column=2, value="[COUNT]")
    row += 1
    ws.cell(row=row, column=1, value="Gaps Identified:")
    ws.cell(row=row, column=2, value="[COUNT]")
    row += 1
    ws.cell(row=row, column=1, value="Critical Findings:")
    ws.cell(row=row, column=2, value="[COUNT]")
    
    # Approval Table
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value="Approval Signatures")
    apply_style(cell, styles["subheader"])
    
    row += 1
    headers = ["Role", "Name", "Signature", "Date", "Comments"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=header)
        apply_style(cell, styles["column_header"])
    
    row += 1
    roles = [
        "Assessment Lead",
        "CISO",
        "IT Operations Manager",
        "Cloud Architect",
        "Information Security Manager"
    ]
    
    for role in roles:
        ws.cell(row=row, column=1, value=role)
        for col in range(1, 6):
            apply_style(ws.cell(row=row, column=col), styles["input_cell"])
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 40


# ============================================================================
# SECTION 13: MAIN EXECUTION
# ============================================================================

def main():
    """Generate the assessment workbook."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.8.31.1 - Environment Architecture Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.31")
    logger.info("=" * 80)
    
    # Create workbook
    logger.info("\nCreating workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    
    # Generate each sheet
    logger.info("Generating Instructions & Legend sheet...")
    create_instructions_sheet(wb, styles)
    
    logger.info("Generating Environment Inventory sheet...")
    create_environment_inventory_sheet(wb, styles)
    
    logger.info("Generating Network Separation sheet...")
    create_network_separation_sheet(wb, styles)
    
    logger.info("Generating Infrastructure Separation sheet...")
    create_infrastructure_separation_sheet(wb, styles)
    
    logger.info("Generating Data Separation sheet...")
    create_data_separation_sheet(wb, styles)
    
    logger.info("Generating Credential Separation sheet...")
    create_credential_separation_sheet(wb, styles)
    
    logger.info("Generating Configuration Consistency sheet...")
    create_configuration_consistency_sheet(wb, styles)
    
    logger.info("Generating Gap Analysis sheet...")
    create_gap_analysis_sheet(wb, styles)
    
    logger.info("Generating Evidence Register sheet...")
    create_evidence_register_sheet(wb, styles)
    
    logger.info("Generating Approval Sign-Off sheet...")
    create_approval_signoff_sheet(wb, styles)
    
    # Save workbook
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.8.31.1_Environment_Architecture_Assessment_{timestamp}.xlsx"
    
    logger.info(f"\nSaving workbook: {filename}")
    wb.save(filename)
    
    logger.info("=" * 80)
    logger.info("✅ SUCCESS!")
    logger.info("=" * 80)
    logger.info(f"\nGenerated: {filename}")
    logger.info("\nNext Steps:")
    logger.info("1. Open the workbook in Excel/LibreOffice Calc")
    logger.info("2. Complete each assessment sheet")
    logger.info("3. Document gaps in Gap_Analysis sheet")
    logger.info("4. Collect evidence for Evidence_Register")
    logger.info("5. Obtain approvals in Approval_Sign_Off sheet")
    logger.info("\n" + "=" * 80)
    

if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
