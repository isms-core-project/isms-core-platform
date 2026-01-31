#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-A.8.2-3-5 - Dashboard Consolidation Utility
================================================================================

ISO/IEC 27001:2022 Controls A.8.2, A.8.3, A.8.5: Authentication & Privileged Access
Operational Utility: Multi-Domain Assessment Dashboard Consolidation

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script consolidates data from five assessment workbooks into the unified
compliance dashboard, providing executive visibility and audit-ready evidence.

**Source Workbooks:**
1. ISMS-IMP-A.8.2-3-5.S1 - Authentication Inventory
2. ISMS-IMP-A.8.2-3-5.S2 - MFA Coverage Assessment
3. ISMS-IMP-A.8.2-3-5.S3 - Privileged Accounts Inventory
4. ISMS-IMP-A.8.2-3-5.S4 - Privileged Access Monitoring
5. ISMS-IMP-A.8.2-3-5.S5 - Access Restrictions Assessment

**Target Dashboard:**
ISMS-IMP-A.8.2-3-5.6 - Compliance Dashboard

**Consolidation Process:**
1. Validate all source workbooks exist
2. Extract gaps from each assessment domain
3. Extract evidence references from each assessment
4. Generate risk entries from critical gaps
5. Generate remediation roadmap from all gaps
6. Populate dashboard sheets with consolidated data
7. Save updated dashboard

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

    python3 consolidate_a8235_dashboard.py

    The script expects source workbooks and dashboard in ../90_workbooks/

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

    - Python 3.8+
    - openpyxl library
    - All 6 workbooks generated and in 90_workbooks folder

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import os
import sys
from datetime import datetime, timedelta
from openpyxl import load_workbook

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


# ==============================================================================
# CONFIGURATION
# ==============================================================================

# Source workbook definitions (filename pattern, assessment area name)
SOURCE_WORKBOOKS = [
    ('ISMS-IMP-A.8.2-3-5.S1_Authentication_Inventory', 'Authentication'),
    ('ISMS-IMP-A.8.2-3-5.S2_MFA_Coverage', 'MFA Coverage'),
    ('ISMS-IMP-A.8.2-3-5.S3_Privileged_Accounts', 'Privileged Accounts'),
    ('ISMS-IMP-A.8.2-3-5.S4_Privileged_Monitoring', 'PAM Monitoring'),
    ('ISMS-IMP-A.8.2-3-5.S5_Access_Restrictions', 'Access Restrictions'),
]

DASHBOARD_PATTERN = 'ISMS-IMP-A.8.2-3-5.6_Compliance_Dashboard'

# Sheets to search for gaps in source workbooks
GAP_SHEET_NAMES = ['Gap_Analysis', 'Gap Analysis', 'MFA_Gaps_Priority', 'Gaps']

# Sheets to search for evidence in source workbooks
EVIDENCE_SHEET_NAMES = ['Evidence_Register', 'Evidence Register', 'Evidence']

# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def find_workbook(directory, pattern):
    """Find workbook matching pattern in directory"""
    for filename in os.listdir(directory):
        if filename.startswith(pattern) and filename.endswith('.xlsx'):
            return os.path.join(directory, filename)
    return None


def safely_write_data(ws, start_row, data):
    """Write data rows to worksheet, return count written"""
    count = 0
    for i, row_data in enumerate(data):
        for j, value in enumerate(row_data):
            if value is not None:
                ws.cell(row=start_row + i, column=j + 1, value=value)
        count += 1
    return count


def extract_gaps_from_workbook(filepath, assessment_area, gap_counter_start=1):
    """Extract gap entries from source workbook"""
    gaps = []
    gap_counter = gap_counter_start

    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)

        # Find gap analysis sheet
        gap_sheet = None
        for sheet_name in GAP_SHEET_NAMES:
            if sheet_name in wb.sheetnames:
                gap_sheet = sheet_name
                break

        if not gap_sheet:
            print(f"    ⚠️  No gap sheet found in {os.path.basename(filepath)}")
            wb.close()
            return gaps, gap_counter

        ws = wb[gap_sheet]

        # Scan for gap data rows (skip header rows, typically start at row 8-10)
        for row in range(8, min(ws.max_row + 1, 100)):
            row_values = [ws.cell(row=row, column=col).value for col in range(1, 12)]

            # Skip empty rows
            if not any(row_values):
                continue

            # Look for gap indicators (status columns with ⚠️, ❌, or text status)
            status = None
            gap_desc = None
            system_name = row_values[0] if row_values[0] else 'Unknown'

            for col_idx, val in enumerate(row_values):
                if val:
                    val_str = str(val)
                    if '⚠️' in val_str or '❌' in val_str or 'Non-Compliant' in val_str or 'Partial' in val_str:
                        status = val_str
                    elif 'gap' in val_str.lower() or len(val_str) > 30:
                        gap_desc = val_str

            if not status and not gap_desc:
                continue

            severity = 'High' if status and ('❌' in str(status) or 'Non-Compliant' in str(status)) else 'Medium'

            gap_entry = [
                f'GAP-{assessment_area[:3].upper()}-{gap_counter:03d}',  # Gap ID
                assessment_area,                                          # Assessment Area
                os.path.basename(filepath),                               # Source Document
                system_name[:50] if system_name else 'Multiple Systems',  # System/Application
                gap_desc[:200] if gap_desc else f'{assessment_area} gap identified',  # Gap Description
                status.replace('⚠️ ', '').replace('❌ ', '')[:30] if status else 'Gap',  # Current State
                'Compliant',                                              # Target State
                severity,                                                 # Severity
                'Open',                                                   # Status
                f'Remediation required for {assessment_area.lower()}',    # Remediation Plan
                None,                                                     # Owner (to be assigned)
                None,                                                     # Target Date
            ]
            gaps.append(gap_entry)
            gap_counter += 1

        wb.close()
        return gaps, gap_counter

    except Exception as e:
        print(f"    ⚠️  Error extracting gaps from {os.path.basename(filepath)}: {e}")
        return gaps, gap_counter


def extract_evidence_from_workbook(filepath, assessment_area):
    """Extract evidence entries from source workbook"""
    evidence = []

    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)

        # Find evidence sheet
        evidence_sheet = None
        for sheet_name in EVIDENCE_SHEET_NAMES:
            if sheet_name in wb.sheetnames:
                evidence_sheet = sheet_name
                break

        if not evidence_sheet:
            wb.close()
            return evidence

        ws = wb[evidence_sheet]

        for row in range(10, min(ws.max_row + 1, 50)):
            evidence_id = ws.cell(row=row, column=1).value
            if evidence_id and str(evidence_id).startswith('EVD-'):
                row_data = [
                    evidence_id,
                    assessment_area,
                    os.path.basename(filepath),
                    None,  # Gap ID link
                    ws.cell(row=row, column=2).value,  # Document Title
                    ws.cell(row=row, column=3).value,  # Document Type
                    ws.cell(row=row, column=4).value,  # Related Control
                    ws.cell(row=row, column=5).value,  # Storage Location
                    ws.cell(row=row, column=6).value,  # Date Collected
                    ws.cell(row=row, column=7).value,  # Collected By
                    ws.cell(row=row, column=8).value,  # Retention Period
                    ws.cell(row=row, column=9).value,  # Status
                ]
                evidence.append(row_data)

        wb.close()
        return evidence

    except Exception as e:
        print(f"    ⚠️  Error extracting evidence from {os.path.basename(filepath)}: {e}")
        return evidence


def generate_risks_from_gaps(gaps):
    """Generate risk entries from critical/high severity gaps"""
    risks = []
    risk_counter = 1

    for gap in gaps:
        gap_id = gap[0]
        area = gap[1]
        system = gap[3]
        gap_desc = gap[4]
        severity = gap[7]

        # Only create risks for High severity gaps
        if severity != 'High':
            continue

        risk_entry = [
            f'RISK-{risk_counter:03d}',                             # Risk ID
            gap_id,                                                  # Gap ID (Link)
            'Security',                                              # Risk Category
            f'Security risk: {gap_desc[:80]}...' if len(gap_desc) > 80 else f'Security risk: {gap_desc}',  # Description
            system,                                                  # Affected System
            'High',                                                  # Likelihood
            'High',                                                  # Impact
            'High',                                                  # Inherent Risk
            'Open',                                                  # Status
            f'Mitigate through remediation of {gap_id}',            # Mitigation Plan
            None,                                                    # Owner
            None,                                                    # Target Date
            None,                                                    # Residual Risk
        ]
        risks.append(risk_entry)
        risk_counter += 1

    return risks


def generate_remediation_from_gaps(gaps):
    """Generate remediation roadmap entries from all gaps"""
    remediation = []
    action_counter = 1
    today = datetime.now()

    for gap in gaps:
        gap_id = gap[0]
        area = gap[1]
        system = gap[3]
        gap_desc = gap[4]
        severity = gap[7]
        remediation_plan = gap[9]

        # Priority based on severity
        if severity == 'High':
            priority = 'Critical'
            days_offset = 30
        elif severity == 'Medium':
            priority = 'High'
            days_offset = 60
        else:
            priority = 'Medium'
            days_offset = 90

        target_date = (today + timedelta(days=days_offset)).strftime('%Y-%m-%d')

        remediation_entry = [
            f'REM-{action_counter:03d}',                             # Action ID
            gap_id,                                                   # Gap ID (Link)
            area,                                                     # Assessment Area
            system,                                                   # System/Application
            remediation_plan if remediation_plan else 'Remediation required',  # Action Required
            priority,                                                 # Priority
            'Planned',                                                # Status
            None,                                                     # Owner
            today.strftime('%Y-%m-%d'),                              # Start Date
            target_date,                                              # Target Date
            None,                                                     # Actual Date
            '0%',                                                     # Progress
            None,                                                     # Notes
        ]
        remediation.append(remediation_entry)
        action_counter += 1

    return remediation


def generate_action_items_from_gaps(gaps):
    """Generate action items from gaps"""
    actions = []
    action_counter = 1
    today = datetime.now()

    for gap in gaps[:50]:  # Limit to 50 action items
        gap_id = gap[0]
        area = gap[1]
        gap_desc = gap[4]
        severity = gap[7]

        priority = 'High' if severity == 'High' else 'Medium'
        due_days = 30 if severity == 'High' else 60
        due_date = (today + timedelta(days=due_days)).strftime('%Y-%m-%d')

        action_entry = [
            f'ACT-{action_counter:03d}',                             # Action ID
            gap_id,                                                   # Related Gap
            f'Address: {gap_desc[:60]}...' if len(gap_desc) > 60 else f'Address: {gap_desc}',  # Action Description
            area,                                                     # Area
            priority,                                                 # Priority
            'Open',                                                   # Status
            None,                                                     # Assigned To
            today.strftime('%Y-%m-%d'),                              # Created Date
            due_date,                                                 # Due Date
            None,                                                     # Completed Date
            None,                                                     # Notes
        ]
        actions.append(action_entry)
        action_counter += 1

    return actions


# ==============================================================================
# MAIN CONSOLIDATION FUNCTION
# ==============================================================================

def consolidate_dashboard():
    """Main consolidation function"""

    print("=" * 80)
    print("ISMS-A.8.2-3-5 DASHBOARD CONSOLIDATION UTILITY")
    print("Authentication & Privileged Access Management")
    print("=" * 80)

    # Determine paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    workbooks_dir = os.path.join(os.path.dirname(script_dir), '90_workbooks')

    print(f"\nWorkbooks directory: {workbooks_dir}")

    if not os.path.exists(workbooks_dir):
        print(f"❌ ERROR: Workbooks directory not found: {workbooks_dir}")
        return False

    # Find dashboard
    dashboard_path = find_workbook(workbooks_dir, DASHBOARD_PATTERN)
    if not dashboard_path:
        print(f"❌ ERROR: Dashboard workbook not found matching: {DASHBOARD_PATTERN}")
        return False

    print(f"Dashboard: {os.path.basename(dashboard_path)}")

    # Find and validate source workbooks
    print("\n[1/5] Validating source workbooks...")
    sources = []
    for pattern, area in SOURCE_WORKBOOKS:
        path = find_workbook(workbooks_dir, pattern)
        if path:
            print(f"  ✓ {area}: {os.path.basename(path)}")
            sources.append((path, area))
        else:
            print(f"  ⚠️  {area}: NOT FOUND ({pattern})")

    if len(sources) < 3:
        print(f"\n❌ ERROR: Insufficient source workbooks ({len(sources)}/5)")
        return False

    # Extract gaps
    print("\n[2/5] Extracting gaps from source workbooks...")
    all_gaps = []
    gap_counter = 1
    for filepath, area in sources:
        gaps, gap_counter = extract_gaps_from_workbook(filepath, area, gap_counter)
        print(f"  • {area}: {len(gaps)} gaps")
        all_gaps.extend(gaps)

    print(f"\n  Total gaps identified: {len(all_gaps)}")

    # Extract evidence
    print("\n[3/5] Extracting evidence from source workbooks...")
    all_evidence = []
    for filepath, area in sources:
        evidence = extract_evidence_from_workbook(filepath, area)
        print(f"  • {area}: {len(evidence)} evidence docs")
        all_evidence.extend(evidence)

    print(f"\n  Total evidence collected: {len(all_evidence)}")

    # Generate derived data
    print("\n[4/5] Generating risks, remediation, and actions...")
    all_risks = generate_risks_from_gaps(all_gaps)
    all_remediation = generate_remediation_from_gaps(all_gaps)
    all_actions = generate_action_items_from_gaps(all_gaps)
    print(f"  • Risks generated: {len(all_risks)}")
    print(f"  • Remediation actions: {len(all_remediation)}")
    print(f"  • Action items: {len(all_actions)}")

    # Populate dashboard
    print("\n[5/5] Writing to dashboard...")
    try:
        wb = load_workbook(dashboard_path)
    except Exception as e:
        print(f"  ❌ Error loading dashboard: {e}")
        return False

    # Populate Gap Analysis sheet
    if 'Gap Analysis' in wb.sheetnames and all_gaps:
        ws = wb['Gap Analysis']
        # Find first empty row after headers (typically row 14+)
        start_row = 14
        for row in range(14, 30):
            if not ws.cell(row=row, column=1).value:
                start_row = row
                break
        count = safely_write_data(ws, start_row, all_gaps)
        print(f"  ✓ Gap Analysis: {count} entries (starting row {start_row})")

    # Populate Risk Register sheet
    if 'Risk Register' in wb.sheetnames and all_risks:
        ws = wb['Risk Register']
        start_row = 21
        for row in range(21, 35):
            if not ws.cell(row=row, column=1).value:
                start_row = row
                break
        count = safely_write_data(ws, start_row, all_risks)
        print(f"  ✓ Risk Register: {count} entries (starting row {start_row})")

    # Populate Remediation Roadmap sheet
    if 'Remediation Roadmap' in wb.sheetnames and all_remediation:
        ws = wb['Remediation Roadmap']
        start_row = 37
        for row in range(37, 50):
            if not ws.cell(row=row, column=1).value:
                start_row = row
                break
        count = safely_write_data(ws, start_row, all_remediation)
        print(f"  ✓ Remediation Roadmap: {count} entries (starting row {start_row})")

    # Populate Evidence Register sheet
    if 'Evidence Register' in wb.sheetnames and all_evidence:
        ws = wb['Evidence Register']
        start_row = 14
        for row in range(14, 25):
            if not ws.cell(row=row, column=1).value:
                start_row = row
                break
        count = safely_write_data(ws, start_row, all_evidence)
        print(f"  ✓ Evidence Register: {count} entries (starting row {start_row})")

    # Populate Action Items sheet
    if 'Action Items & Follow-up' in wb.sheetnames and all_actions:
        ws = wb['Action Items & Follow-up']
        start_row = 14
        for row in range(14, 25):
            if not ws.cell(row=row, column=1).value:
                start_row = row
                break
        count = safely_write_data(ws, start_row, all_actions)
        print(f"  ✓ Action Items: {count} entries (starting row {start_row})")

    # Save dashboard
    try:
        wb.save(dashboard_path)
        print(f"\n💾 Saved: {os.path.basename(dashboard_path)}")
    except Exception as e:
        print(f"  ❌ Error saving: {e}")
        wb.close()
        return False

    wb.close()

    # Summary
    print("\n" + "=" * 80)
    print("✅ CONSOLIDATION COMPLETE")
    print("=" * 80)
    print(f"\nSummary:")
    print(f"  • Source workbooks processed: {len(sources)}/5")
    print(f"  • Gap Analysis: {len(all_gaps)} gaps consolidated")
    print(f"  • Risk Register: {len(all_risks)} risks documented")
    print(f"  • Remediation Roadmap: {len(all_remediation)} actions planned")
    print(f"  • Evidence Register: {len(all_evidence)} evidence documents")
    print(f"  • Action Items: {len(all_actions)} follow-up items")
    total_data = len(all_gaps) + len(all_risks) + len(all_remediation) + len(all_evidence) + len(all_actions)
    print(f"\n  Total Data Points: {total_data}")
    print("\n🎯 Dashboard is now ready for executive review!")
    print("=" * 80 + "\n")

    return True


# ==============================================================================
# ENTRY POINT
# ==============================================================================

if __name__ == "__main__":
    success = consolidate_dashboard()
    sys.exit(0 if success else 1)

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (syntax validated, structure verified)
# QA_TOOL: Claude Code Deep Scan
# QA_NOTE: STANDARDIZATION - License header, logging, main() pattern applied
# =============================================================================
