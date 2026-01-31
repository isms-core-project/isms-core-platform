#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.2-3-5.S4 - Privileged Access Monitoring Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.2, A.8.3, A.8.5: Authentication & Privileged Access
Assessment Domain 4 of 6: Privileged Access Activity Monitoring

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific monitoring capabilities, SIEM deployment, and
privileged access oversight requirements.

Key customization areas:
1. Session recording scope (CyberArk PSM, Teleport per your deployment)
2. SIEM integration (Splunk, Sentinel per your platform)
3. Anomaly detection thresholds (based on your baseline)
4. Alert escalation procedures (per your SOC operations)
5. Review frequency requirements (aligned with your governance)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for authentication)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

**Purpose:**
Enables systematic assessment of privileged access monitoring, session
recording, and anomaly detection capabilities, supporting ISO 27001:2022
Control A.8.2 privileged access oversight requirements.

**Assessment Scope:**
- Privileged access activity logging
- Privileged session recording coverage
- Privileged command execution logging (sudo, runas)
- Off-hours privileged access detection
- Unusual privileged activity identification
- Failed privileged access attempts
- SIEM integration for privileged access alerts
- Privileged access review completion
- Access anomaly investigation status
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Privileged monitoring assessment guidance
2. Privileged Activity Log Coverage - Logging implementation status
3. Session Recording Status - Recorded vs. non-recorded privileged sessions
4. Privileged Command Logging - Sudo/runas execution tracking
5. Access Anomalies - Off-hours access, unusual patterns
6. Failed Access Attempts - Authentication failures, access denials
7. SIEM Integration - Alert configuration and escalation
8. Access Review Results - Periodic review outcomes
9. Investigation Status - Anomaly resolution tracking
10. Evidence Register - Audit evidence tracking
11. Approval & Sign-Off - Stakeholder review workflow

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a8235_4_privileged_monitoring.py

Requirements:
    - Python 3.8+
    - openpyxl library: pip install openpyxl

Output:
    ISMS-IMP-A.8.2-3-5.S4_Privileged_Monitoring_Assessment_YYYYMMDD.xlsx
"""
# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# ============================================================================
# SECTION 1: CONSTANTS AND CONFIGURATION
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
KEY = '\U0001F511'    # 🔑 Key
PACKAGE = '\U0001F4E6' # 📦 Package
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow
WORKBOOK_NAME = "Privileged Access Monitoring Assessment"
DOCUMENT_ID = "ISMS-IMP-A.8.2-3-5.S4"
VERSION = "1.0"
CONTROL_REF = "ISO/IEC 27001:2022 - Controls A.8.2, A.8.3, A.8.5: Authentication & Privileged Access"
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
# Workbook structure
SESSION_RECORDING_COUNT = 150    # Session recording coverage tracking
COMMAND_LOGGING_COUNT = 100      # Command logging verification
ANOMALY_RULES_COUNT = 30         # Alert rule definitions
ACCESS_ACTIVITY_COUNT = 200      # Historical access log entries
ALERT_TRACKING_COUNT = 100       # Security alert response tracking
OFF_HOURS_COUNT = 80             # Off-hours access entries
FAILED_LOGIN_COUNT = 100         # Failed login attempts
TIER_VIOLATION_COUNT = 50        # Tier isolation violations
REVIEW_STATUS_COUNT = 40         # Quarterly review status
EVIDENCE_ROW_COUNT = 30          # Evidence register entries
# Session recording status
RECORDING_STATUS = [
    f"{CHECK} Enabled - Active Recording",
    f"{CHECK} Enabled - Verified",
    f"{WARNING} Enabled - Storage Issues",
    f"{XMARK} Not Enabled (Gap)",
    "🔄 Implementation In Progress",
    "➖ N/A (Not Applicable)"
]
# Command logging types
COMMAND_LOGGING_TYPES = [
    "PowerShell Script Block Logging",
    "sudo Command Logging (Linux)",
    "Windows Security Event Logging",
    "Database DDL/DML Logging",
    "Cloud API Audit Logging (CloudTrail/Activity Log)",
    "PAM Session Command Logging",
    "Network Device Command Logging",
    "Application Admin Action Logging"
]

# Anomaly types
ANOMALY_TYPES = [
    "🔴 Off-Hours Access (Outside Business Hours)",
    "🔴 Unusual Location / Impossible Travel",
    "🔴 Tier Isolation Violation",
    "🟠 Failed Login Attempts (>3 in 5 min)",
    "🟠 Long Session Duration (>8 hours)",
    "🟠 Break-Glass Account Usage",
    "🟡 Unusual Command Pattern",
    "🟡 Multiple Concurrent Sessions"
]
# Alert severity
ALERT_SEVERITY = [
    "🔴 Critical",
    "🟠 High",
    "🟡 Medium",
    "🟢 Low",
    "ℹ️ Informational"
]
# Alert status
ALERT_STATUS = [
    "🔴 Open - Investigating",
    "🟡 Open - Awaiting Response",
    "🟢 Closed - Legitimate Activity",
    "🔴 Closed - Security Incident",
    "⚪ Closed - False Positive"
]
# Compliance status
COMPLIANCE_STATUS = [
    f"{CHECK} Compliant",
    f"{WARNING} Partial Compliance",
    f"{XMARK} Non-Compliant",
    "🔄 In Progress",
    "📋 Under Review",
    "❓ Unknown",
    "➖ N/A"
]
# Admin tiers
ADMIN_TIERS = [
    "Tier 0 (Domain/Enterprise/Critical)",
    "Tier 1 (Server/Application)",
    "Tier 2 (Workstation/Endpoint)",
    "N/A (Non-Privileged)"
]
# Review status
REVIEW_STATUS = [
    f"{CHECK} Completed On Time",
    f"{WARNING} Completed Late",
    f"{XMARK} Overdue",
    "📋 Scheduled"
]
# SECTION 2: STYLE DEFINITIONS
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    return {
        'header': {
            'font': {'name': 'Calibri', 'size': 11, 'bold': True, 'color': 'FFFFFF'},
            'fill': {'patternType': 'solid', 'fgColor': '003366'},
            'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'subheader': {
            'font': {'name': 'Calibri', 'size': 11, 'bold': True, 'color': '000000'},
            'fill': {'patternType': 'solid', 'fgColor': 'D8E4F8'},
            'alignment': {'horizontal': 'left', 'vertical': 'center', 'wrap_text': True},
        },
        'data': {
            'font': {'name': 'Calibri', 'size': 10, 'color': '000000'},
            'fill': {'patternType': 'solid', 'fgColor': 'FFFFFF'},
            'alignment': {'horizontal': 'left', 'vertical': 'center', 'wrap_text': False},
        },
        'title': {
            'font': {'name': 'Calibri', 'size': 16, 'bold': True, 'color': '003366'},
            'alignment': {'horizontal': 'left', 'vertical': 'center'}
        },
        'critical_alert': {
            'font': {'name': 'Calibri', 'size': 10, 'bold': True, 'color': 'C00000'},
            'fill': {'patternType': 'solid', 'fgColor': 'FCE4D6'},
            'alignment': {'horizontal': 'left', 'vertical': 'center'},
        }
    }
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell with NEW objects."""
    if 'font' in style_dict:
        cell.font = Font(**style_dict['font'])
    if 'fill' in style_dict:
        cell.fill = PatternFill(**style_dict['fill'])
    if 'alignment' in style_dict:
        cell.alignment = Alignment(**style_dict['alignment'])
    if 'border' in style_dict:
        cell.border = Border(
            left=Side(style=style_dict['border'].get('left', 'thin')),
            right=Side(style=style_dict['border'].get('right', 'thin')),
            top=Side(style=style_dict['border'].get('top', 'thin')),
            bottom=Side(style=style_dict['border'].get('bottom', 'thin'))
        )
# SECTION 3: WORKBOOK CREATION
def create_workbook():
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    # Create sheets matching specification
    sheets = [
        "Instructions_Legend",
        "Session_Recording_Coverage",
        "Privileged_Command_Logging",
        "Anomaly_Detection_Rules",
        "Privileged_Access_Activity",
        "Alert_Response_Tracking",
        "Off_Hours_Access_Log",
        "Failed_Login_Analysis",
        "Tier_Violation_Incidents",
        "Quarterly_Review_Status",
        "Evidence_Register",
        "Approval_Sign_Off"
    ]
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    return wb
# SECTION 4: SHEET 1 - INSTRUCTIONS & LEGEND
def populate_instructions(wb, styles):
    """Populate Instructions & Legend sheet."""
    ws = wb["Instructions_Legend"]
    # Title
    ws['A1'] = f"{DOCUMENT_ID}  -  Privileged Access Monitoring & Anomaly Detection Assessment\n{CONTROL_REF}"
    apply_style(ws['A1'], styles['title'])
    ws.row_dimensions[1].height = 40
    # Document metadata (standardized rows 3-6)
    ws['A3'] = 'Document ID:'
    ws['A3'].font = Font(bold=True)
    ws['B3'] = DOCUMENT_ID
    ws['A4'] = 'Assessment:'
    ws['A4'].font = Font(bold=True)
    ws['B4'] = 'Privileged Access Monitoring'
    ws['A5'] = 'Version:'
    ws['A5'].font = Font(bold=True)
    ws['B5'] = VERSION
    ws['A6'] = 'Generated:'
    ws['A6'].font = Font(bold=True)
    ws['B6'] = datetime.now().strftime('%d.%m.%Y %H:%M')
    ws.column_dimensions['B'].width = 40
    # Monitoring Imperative
    row = 8
    ws[f'A{row}'] = f"{WARNING} MONITORING = DETECTION CAPABILITY"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='C00000')
    ws[f'A{row}'].fill = PatternFill(patternType='solid', fgColor='FCE4D6')
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    ws[f'A{row}'] = "Without monitoring, privileged access abuse is invisible. Session recording provides evidence of what privileged users did. Real-time anomaly detection enables rapid incident response. Monitoring is not optional for privileged access."
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 45
    row += 2
    ws[f'A{row}'] = "MONITORING COVERAGE REQUIREMENTS"
    apply_style(ws[f'A{row}'], styles['subheader'])
    requirements = [
        ("Session Recording", "Tier 0: 100% | Tier 1 Production: 90%+ | Tier 2: Optional", "🔴 Critical"),
        ("Command Logging", "All privileged commands (sudo, PowerShell, database DDL, cloud APIs)", "🔴 Critical"),
        ("Real-Time Alerts", "Off-hours access, failed logins, tier violations, unusual activity", "🔴 Critical"),
        ("Response Time", "Critical alerts: <30 min average | High alerts: <2 hours", "🟠 High"),
        ("Log Retention", "90 days online, 1 year archived minimum", "🟡 Medium"),
        ("SIEM Integration", "Forward privileged access logs to SIEM (A.8.16)", "🟠 High")
    ]
    row += 1
    ws[f'A{row}'] = "Monitoring Component"
    ws[f'B{row}'] = "Requirement"
    ws[f'C{row}'] = "Priority"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f'{col}{row}'], styles['header'])
    for component, requirement, priority in requirements:
        row += 1
        ws[f'A{row}'] = component
        ws[f'B{row}'] = requirement
        ws[f'C{row}'] = priority
        apply_style(ws[f'A{row}'], styles['data'])
        apply_style(ws[f'B{row}'], styles['data'])
        apply_style(ws[f'C{row}'], styles['data'])
    ws[f'A{row}'] = "SESSION RECORDING TARGETS"
    recording_targets = [
        ("Tier 0 Sessions", "100%", "MANDATORY - All domain controller, cloud global admin, critical infrastructure access"),
        ("Tier 1 Production", "90%+", "HIGHLY RECOMMENDED - Production server and database access"),
        ("Tier 1 Non-Production", "50%+", "RECOMMENDED - Test/dev server access (less critical)"),
        ("Tier 2 Sessions", "Optional", "OPTIONAL - Workstation local admin (low risk)")
    ]
    row += 1
    ws[f'A{row}'] = "Tier / Environment"
    ws[f'B{row}'] = "Target Coverage"
    ws[f'C{row}'] = "Justification"
    for tier, target, justification in recording_targets:
        ws[f'A{row}'] = tier
        ws[f'B{row}'] = target
        ws[f'C{row}'] = justification
    ws[f'A{row}'] = "ANOMALY DETECTION & ALERT FRAMEWORK"
    alerts = [
        ("🔴 CRITICAL", "Off-Hours Access (Tier 0)", "Alert immediately, investigate within 30 min", "Tier 0 access outside business hours unusual"),
        ("🔴 CRITICAL", "Tier Isolation Violation", "Alert immediately, terminate session", "Tier 0 account on Tier 1/2 system = security breach"),
        ("🔴 CRITICAL", "Break-Glass Usage", "Alert immediately, executive notification", "Sealed credentials used = emergency"),
        ("🟠 HIGH", "Failed Login (>3 in 5 min)", "Alert within 1 hour, investigate", "Brute force attack indicator"),
        ("🟠 HIGH", "Unusual Location / Impossible Travel", "Alert within 1 hour, verify user", "Account compromise indicator"),
        ("🟡 MEDIUM", "Long Session (>8 hours)", "Daily report, review next day", "Session management issue"),
        ("🟡 MEDIUM", "Off-Hours Access (Tier 1/2)", "Daily report, review next day", "Standard users less critical")
    ]
    row += 1
    ws[f'A{row}'] = "Severity"
    ws[f'B{row}'] = "Alert Condition"
    ws[f'C{row}'] = "Response Requirement"
    ws[f'D{row}'] = "Rationale"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['header'])
    for severity, condition, response, rationale in alerts:
        ws[f'A{row}'] = severity
        ws[f'B{row}'] = condition
        ws[f'C{row}'] = response
        ws[f'D{row}'] = rationale
        apply_style(ws[f'D{row}'], styles['data'])
    ws[f'A{row}'] = "COMMAND LOGGING REQUIREMENTS"
    command_logging = [
        f"{BULLET} PowerShell: Script Block Logging enabled (captures all PowerShell commands)",
        f"{BULLET} Linux: sudo command logging (/var/log/auth.log), auditd for privileged actions",
        f"{BULLET} Windows: Security Event ID 4688 (process creation), 4672 (privileged logon)",
        f"{BULLET} Database: DDL logging (CREATE, ALTER, DROP), DML on sensitive tables",
        f"{BULLET} Cloud: CloudTrail (AWS), Activity Log (Azure), Audit Logs (GCP)",
        f"{BULLET} Network Devices: TACACS+ command accounting",
        f"{BULLET} PAM: Session command capture (keystroke or video)",
        f"{BULLET} Applications: Admin action logging (user creation, permission changes)"
    ]
    for log_req in command_logging:
        ws[f'A{row}'] = log_req
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = 22
    ws[f'A{row}'] = f"{TARGET} REMEMBER: 'We monitor privileged access' means nothing without actual session recordings and real-time alerts."
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, italic=True, color='C00000')
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[row].height = 30
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
# SECTION 5: SHEET 2 - SESSION RECORDING COVERAGE
def populate_session_recording(wb, styles):
    """Populate Session Recording Coverage sheet."""
    ws = wb["Session_Recording_Coverage"]
    ws['A1'] = "Session Recording Coverage Assessment"
    ws.merge_cells('A1:L1')
    ws['A2'] = "Track session recording deployment (Target: Tier 0 100%, Tier 1 Production 90%+)"
    ws.merge_cells('A2:L2')
    # Headers
    headers = [
        "Account / System",
        "Admin Tier",
        "Environment",
        "Recording Status",
        "Recording Method",
        "Storage Location",
        "Retention Period",
        "Last Recording Date",
        "Playback Tested",
        "Compliance Target",
        "Compliance Status",
        "Notes"
    ]
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    # Pre-format data rows
    for row_num in range(5, 5 + SESSION_RECORDING_COUNT):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    # Add dropdowns
    # Column B: Admin Tier
    dv_tier = DataValidation(type="list", formula1=f'"{",".join(ADMIN_TIERS)}"', allow_blank=True)
    ws.add_data_validation(dv_tier)
    dv_tier.add(f'B5:B{4 + SESSION_RECORDING_COUNT}')
    # Column C: Environment
    dv_env = DataValidation(type="list", formula1='"Production,Non-Production,Test,Development"', allow_blank=True)
    ws.add_data_validation(dv_env)
    dv_env.add(f'C5:C{4 + SESSION_RECORDING_COUNT}')
    # Column D: Recording Status
    dv_status = DataValidation(type="list", formula1=f'"{",".join(RECORDING_STATUS)}"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f'D5:D{4 + SESSION_RECORDING_COUNT}')
    # Column E: Recording Method
    dv_method = DataValidation(type="list", formula1='"Video Recording,Keystroke Logging,Both (Video + Keystroke),None"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add(f'E5:E{4 + SESSION_RECORDING_COUNT}')
    # Column I: Playback Tested
    dv_tested = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_tested)
    dv_tested.add(f'I5:I{4 + SESSION_RECORDING_COUNT}')
    # Column J: Compliance Target
    dv_target = DataValidation(type="list", formula1='"Tier 0: 100%,Tier 1 Prod: 90%+,Tier 1 Non-Prod: 50%+,Optional"', allow_blank=True)
    ws.add_data_validation(dv_target)
    dv_target.add(f'J5:J{4 + SESSION_RECORDING_COUNT}')
    # Column K: Compliance Status
    dv_compliance = DataValidation(type="list", formula1=f'"{",".join(COMPLIANCE_STATUS)}"', allow_blank=True)
    ws.add_data_validation(dv_compliance)
    dv_compliance.add(f'K5:K{4 + SESSION_RECORDING_COUNT}')
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 35
    # Freeze panes
    ws.freeze_panes = 'A5'
    # Summary metrics
    row = 5 + SESSION_RECORDING_COUNT + 2
    ws[f'A{row}'] = "RECORDING COVERAGE SUMMARY"
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    ws[f'A{row}'] = "Tier 0 Recording Coverage:"
    ws[f'B{row}'] = f'=COUNTIFS(B5:B{4+SESSION_RECORDING_COUNT},"Tier 0*",D5:D{4+SESSION_RECORDING_COUNT},"{CHECK} Enabled*")/COUNTIF(B5:B{4+SESSION_RECORDING_COUNT},"Tier 0*")'
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'C{row}'] = "Target: 100%"
    row += 1
    ws[f'A{row}'] = "Tier 1 Prod Recording Coverage:"
    ws[f'B{row}'] = f'=COUNTIFS(B5:B{4+SESSION_RECORDING_COUNT},"Tier 1*",C5:C{4+SESSION_RECORDING_COUNT},"Production",D5:D{4+SESSION_RECORDING_COUNT},"{CHECK} Enabled*")/COUNTIFS(B5:B{4+SESSION_RECORDING_COUNT},"Tier 1*",C5:C{4+SESSION_RECORDING_COUNT},"Production")'
    ws[f'C{row}'] = "Target: 90%+"
# SECTION 6: REMAINING SHEETS
def populate_remaining_sheets(wb, styles):
    """Populate remaining sheets with standard structure."""
    header_row = 4  # Standard header row for all sheets

    # Sheet 3: Privileged Command Logging
    ws_logging = wb["Privileged_Command_Logging"]
    ws_logging['A1'] = "Privileged Command Logging Verification"
    apply_style(ws_logging['A1'], styles['title'])
    ws_logging['A2'] = "Verify command logging is enabled for all privileged activities"
    logging_headers = ["System / Platform", "Admin Tier", "Logging Type", "Status", "Log Destination", "Retention", "SIEM Integrated", "Compliance"]
    for col_idx, header in enumerate(logging_headers, start=1):
        cell = ws_logging.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + COMMAND_LOGGING_COUNT):
        for col_idx in range(1, len(logging_headers) + 1):
            cell = ws_logging.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])

    # Sheet 4: Anomaly Detection Rules
    ws_anomaly = wb["Anomaly_Detection_Rules"]
    ws_anomaly['A1'] = "Anomaly Detection & Alert Rules"
    apply_style(ws_anomaly['A1'], styles['title'])
    anomaly_headers = ["Alert Rule Name", "Anomaly Type", "Severity", "Detection Logic", "Alert Destination", "Response SLA", "Status"]
    for col_idx, header in enumerate(anomaly_headers, start=1):
        cell = ws_anomaly.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + ANOMALY_RULES_COUNT):
        for col_idx in range(1, len(anomaly_headers) + 1):
            cell = ws_anomaly.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])

    # Sheet 5: Privileged Access Activity
    ws_activity = wb["Privileged_Access_Activity"]
    ws_activity['A1'] = "Privileged Access Activity Log"
    apply_style(ws_activity['A1'], styles['title'])
    activity_headers = ["Date/Time", "Account", "User", "System", "Action", "Duration", "Location", "Anomaly Detected", "Review Status"]
    for col_idx, header in enumerate(activity_headers, start=1):
        cell = ws_activity.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + ACCESS_ACTIVITY_COUNT):
        for col_idx in range(1, len(activity_headers) + 1):
            cell = ws_activity.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])

    # Sheet 6: Alert Response Tracking
    ws_alerts = wb["Alert_Response_Tracking"]
    ws_alerts['A1'] = "Security Alert Response Tracking"
    apply_style(ws_alerts['A1'], styles['title'])
    alert_headers = ["Alert ID", "Alert Time", "Alert Type", "Severity", "Account", "Assigned To", "Response Time (min)", "Resolution", "Status"]
    for col_idx, header in enumerate(alert_headers, start=1):
        cell = ws_alerts.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + ALERT_TRACKING_COUNT):
        for col_idx in range(1, len(alert_headers) + 1):
            cell = ws_alerts.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])

    # Add dropdown for Status
    dv_alert_status = DataValidation(type="list", formula1=f'"{",".join(ALERT_STATUS)}"', allow_blank=True)
    ws_alerts.add_data_validation(dv_alert_status)
    dv_alert_status.add(f'I5:I{4 + ALERT_TRACKING_COUNT}')

    # Sheet 7: Off-Hours Access Log
    ws_offhours = wb["Off_Hours_Access_Log"]
    ws_offhours['A1'] = "Off-Hours Privileged Access Log"
    apply_style(ws_offhours['A1'], styles['title'])
    ws_offhours['A2'] = "Review privileged access outside business hours (weekdays 6PM-8AM, weekends, holidays)"
    offhours_headers = ["Date/Time", "Account", "User", "System", "Tier", "Business Justification", "Approved By", "Review Status"]
    for col_idx, header in enumerate(offhours_headers, start=1):
        cell = ws_offhours.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + OFF_HOURS_COUNT):
        for col_idx in range(1, len(offhours_headers) + 1):
            cell = ws_offhours.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])

    # Sheet 8: Failed Login Analysis
    ws_failed = wb["Failed_Login_Analysis"]
    ws_failed['A1'] = "Failed Privileged Login Analysis"
    apply_style(ws_failed['A1'], styles['title'])
    ws_failed['A2'] = "Track failed authentication attempts (>3 failures in 5 min = alert)"
    failed_headers = ["Date/Time", "Account", "Source IP", "System", "Failure Count", "Time Window", "Alert Generated", "Investigation"]
    for col_idx, header in enumerate(failed_headers, start=1):
        cell = ws_failed.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + FAILED_LOGIN_COUNT):
        for col_idx in range(1, len(failed_headers) + 1):
            cell = ws_failed.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])

    # Sheet 9: Tier Violation Incidents
    ws_tier = wb["Tier_Violation_Incidents"]
    ws_tier['A1'] = "Tier Isolation Violation Tracking (Should be ZERO)"
    apply_style(ws_tier['A1'], styles['title'])
    tier_headers = ["Date", "Account", "Account Tier", "System", "System Tier", "Violation Type", "Action Taken", "Root Cause", "Status"]
    for col_idx, header in enumerate(tier_headers, start=1):
        cell = ws_tier.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + TIER_VIOLATION_COUNT):
        for col_idx in range(1, len(tier_headers) + 1):
            cell = ws_tier.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])

    # Sheet 10: Quarterly Review Status
    ws_review = wb["Quarterly_Review_Status"]
    ws_review['A1'] = "Quarterly Privileged Access Review Status"
    apply_style(ws_review['A1'], styles['title'])
    review_headers = ["Quarter", "Review Due Date", "Privileged Users Reviewed", "Completion Date", "Days Late/Early", "Review Status", "Findings"]
    for col_idx, header in enumerate(review_headers, start=1):
        cell = ws_review.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + REVIEW_STATUS_COUNT):
        for col_idx in range(1, len(review_headers) + 1):
            cell = ws_review.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
        # Formula for Days Late/Early
        e_cell = ws_review.cell(row=row_num, column=5)
        e_cell.value = f'=IF(AND(B{row_num}<>"",D{row_num}<>""),D{row_num}-B{row_num},"")'

    # Add dropdown for Review Status
    dv_review_status = DataValidation(type="list", formula1=f'"{",".join(REVIEW_STATUS)}"', allow_blank=True)
    ws_review.add_data_validation(dv_review_status)
    dv_review_status.add(f'F5:F{4 + REVIEW_STATUS_COUNT}')

    # Sheet 11: Evidence Register
    ws_evidence = wb["Evidence_Register"]
    ws_evidence['A1'] = "Privileged Monitoring Evidence Register"
    apply_style(ws_evidence['A1'], styles['title'])
    evidence_headers = ["Evidence ID", "Evidence Type", "Description", "Source", "Date", "Collected By", "Verification"]
    for col_idx, header in enumerate(evidence_headers, start=1):
        cell = ws_evidence.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + EVIDENCE_ROW_COUNT):
        for col_idx in range(1, len(evidence_headers) + 1):
            cell = ws_evidence.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    # Sheet 12: Approval Sign-Off
    ws_approval = wb["Approval_Sign_Off"]
    ws_approval['A1'] = "Privileged Monitoring Assessment Approval"
    apply_style(ws_approval['A1'], styles['title'])
    ws_approval['A3'] = "Assessment Completion"
    apply_style(ws_approval['A3'], styles['subheader'])
    ws_approval['A4'] = "Assessment Date:"
    ws_approval['A5'] = "Assessed By:"
    ws_approval['A7'] = "Approval Workflow"
    apply_style(ws_approval['A7'], styles['subheader'])
    approval_rows = [
        ("Level 1", "Security Analyst", "", "", ""),
        ("Level 2", "Security Architect", "", "", ""),
        ("Level 3", "CISO", "", "", "")
    ]
    row = 8
    ws_approval[f'A{row}'] = "Level"
    ws_approval[f'B{row}'] = "Role"
    ws_approval[f'C{row}'] = "Name"
    ws_approval[f'D{row}'] = "Date"
    ws_approval[f'E{row}'] = "Signature"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws_approval[f'{col}{row}'], styles['header'])
    for level, role, name, date, sig in approval_rows:
        row += 1
        ws_approval[f'A{row}'] = level
        ws_approval[f'B{row}'] = role
        ws_approval[f'C{row}'] = name
        ws_approval[f'D{row}'] = date
        ws_approval[f'E{row}'] = sig
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws_approval[f'{col}{row}'], styles['data'])

# SECTION 7: MAIN GENERATION FUNCTION
def generate_workbook():
    """Main function to generate complete workbook."""
    logger.info("")
    logger.info("╔════════════════════════════════════════════════════════════════╗")
    logger.info("║  ISMS Assessment A.8.2/3/5 - Authentication & PAM Framework    ║")
    logger.info("║  Workbook 4: Privileged Access Monitoring Assessment           ║")
    logger.info("║                                                                ║")
    logger.info("║  'Without monitoring, abuse is invisible'                      ║")
    logger.info("╚════════════════════════════════════════════════════════════════╝")
    logger.info("Creating workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    logger.info("Populating Instructions & Legend...")
    populate_instructions(wb, styles)
    logger.info("Populating Session Recording Coverage...")
    populate_session_recording(wb, styles)
    logger.info("Populating remaining sheets...")
    populate_remaining_sheets(wb, styles)
    # Save workbook
    filename = f"ISMS-IMP-A.8.2-3-5.S4_Privileged_Monitoring_{GENERATED_TIMESTAMP}.xlsx"
    wb.save(filename)
    logger.info("=" * 70)
    logger.info("{CHECK} Workbook generated successfully: {filename}")
    logger.info("Next Steps:")
    logger.info("  1. Verify session recording coverage (Tier 0: 100% target)")
    logger.info("  2. Confirm command logging enabled (PowerShell, sudo, database, cloud)")
    logger.info("  3. Configure anomaly detection rules in SIEM")
    logger.info("  4. Track alert response times (<30 min for critical)")
    logger.info("  5. Review off-hours access log weekly")
    logger.error("  6. Investigate failed login patterns")
    logger.info("  7. Monitor tier violation incidents (target: ZERO)")
    logger.info("  8. Complete quarterly privileged access reviews")
    logger.info("  9. Collect evidence in Evidence_Register")
    logger.info(" 10. Obtain approvals in Approval_Sign_Off")
    logger.info("Monitoring Targets:")
    logger.info("  • Session Recording: Tier 0 100%, Tier 1 Prod 90%+")
    logger.info("  • Critical Alert Response: <30 minutes average")
    logger.info("  • Tier Violations: ZERO (immediate investigation)")
    logger.info("  • Quarterly Reviews: 100% completion on time")
    return filename
def validate_workbook(filename):
    """Validate generated workbook."""
    logger.info("Running validation...")
    logger.info("-" * 70)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        expected_sheets = [
            "Instructions_Legend",
            "Session_Recording_Coverage",
            "Privileged_Command_Logging",
            "Anomaly_Detection_Rules",
            "Privileged_Access_Activity",
            "Alert_Response_Tracking",
            "Off_Hours_Access_Log",
            "Failed_Login_Analysis",
            "Tier_Violation_Incidents",
            "Quarterly_Review_Status",
            "Evidence_Register",
            "Approval_Sign_Off"
        ]
        # Check sheet count
        if len(wb.sheetnames) != 12:
            logger.info(f"  ❌ Expected 12 sheets, found {len(wb.sheetnames)}")
            return False
        logger.info(f"  ✅ Sheet count: {len(wb.sheetnames)}")
        # Check sheet names
        for sheet in expected_sheets:
            if sheet in wb.sheetnames:
                logger.info(f"  ✅ Found: {sheet}")
            else:
                logger.info(f"  ❌ Missing: {sheet}")
                return False
        logger.info("")
        logger.info("Validation Result: ✅ PASSED")
        wb.close()
        return True
    except Exception as e:
        logger.error(f"  ❌ Validation error: {str(e)}")
        return False
# SECTION 8: ENTRY POINT
if __name__ == "__main__":
    try:
        # Generate the workbook
        output_file = generate_workbook()
        # Validate the output
        if validate_workbook(output_file):
            logger.info("{CHECK} Successfully generated: {output_file}")
            sys.exit(0)
        else:
            logger.info("{WARNING} Validation warnings - please review the output")
            sys.exit(1)
    except ImportError as e:
        logger.error("ERROR: Missing required library")
        logger.info("-" * 70)
        logger.info(f"  {str(e)}")
        logger.info("Please install openpyxl:")
        logger.info("  pip install openpyxl")
        sys.exit(1)
    except Exception as e:
        logger.error("ERROR: Generation failed")
        import traceback
        traceback.print_exc()
        sys.exit(1)
# =============================================================================
# END OF GENERATOR SCRIPT
#
# Document Control:
#   Version: 1.0
#   Created: 2025-01-11
#   Author:               [Organization] ISMS Implementation Team
# Change History:
#   1.0 - Initial version
#       - 12 sheets for comprehensive privileged access monitoring
#       - Session recording coverage tracking (Tier 0 100%, Tier 1 Prod 90%+)
#       - Command logging verification (PowerShell, sudo, database, cloud)
#       - Anomaly detection rules and alert framework
#       - Real-time alert response tracking (<30 min critical)
#       - Off-hours access review
#       - Failed login analysis
#       - Tier isolation violation tracking (target: ZERO)
#       - Quarterly review status tracking
# Dependencies:
#   - Python 3.7+
#   - openpyxl >= 3.0.0
# Output:
#   Privileged_Monitoring_Assessment_YYYYMMDD.xlsx

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
