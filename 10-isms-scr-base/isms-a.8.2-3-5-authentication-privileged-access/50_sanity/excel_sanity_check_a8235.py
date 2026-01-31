#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
Excel Workbook Sanity Checker - ISMS A.8.2/3/5 Assessment Workbooks
================================================================================

Diagnostic utility for troubleshooting Excel's "file level validation and repair"
warnings when opening A.8.2/3/5 authentication and privileged access assessment
workbooks.

**Purpose:**
Identifies common openpyxl-generated Excel issues that trigger repair warnings:
- Formula syntax errors and invalid sheet references
- Data validation conflicts and overlapping ranges
- Style attribute inconsistencies
- Merged cell content issues
- Worksheet structure problems

**When to Use:**
- Excel displays repair warnings when opening generated workbooks
- After modifying assessment generator scripts
- Before distributing workbooks to stakeholders
- Quality assurance validation before consolidation

**Usage:**
    python3 excel_sanity_check_a8235.py ISMS_A_8_2_3_5_X_Assessment_YYYYMMDD.xlsx
    
    Works with any A.8.2/3/5 assessment workbook (domains 1-6)

**Output:**
- Diagnostic report with issue categorization (Critical/Warning)
- Recommended remediation actions
- Structural health summary

Control Reference: ISO/IEC 27001:2022 Annex A Controls A.8.2, A.8.3, A.8.5
Script Type: Quality Assurance Utility
Version: 1.0

Usage:
    python3 excel_sanity_check_a8235.py Authentication_Inventory_Assessment_20250111.xlsx
    python3 excel_sanity_check_a8235.py MFA_Coverage_Assessment_20250111.xlsx
    python3 excel_sanity_check_a8235.py Privileged_Accounts_Assessment_20250111.xlsx
    python3 excel_sanity_check_a8235.py Privileged_Monitoring_Assessment_20250111.xlsx
    python3 excel_sanity_check_a8235.py Access_Restrictions_Assessment_20250111.xlsx
    python3 excel_sanity_check_a8235.py Authentication_PAM_Dashboard_20250111.xlsx

"""

import sys
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ============================================================================
# WORKBOOK TYPE DETECTION
# ============================================================================

def detect_workbook_type(filename):
    """Detect which A.8.2/3/5 workbook type this is based on filename."""
    filename_lower = filename.lower()
    
    if 'authentication_inventory' in filename_lower or 'wb1' in filename_lower:
        return 'WB1', 'Authentication Inventory & Methods (A.8.5)'
    elif 'mfa_coverage' in filename_lower or 'wb2' in filename_lower:
        return 'WB2', 'MFA Coverage Assessment (A.8.5)'
    elif 'privileged_accounts' in filename_lower or 'wb3' in filename_lower:
        return 'WB3', 'Privileged Accounts Inventory (A.8.2)'
    elif 'privileged_monitoring' in filename_lower or 'wb4' in filename_lower:
        return 'WB4', 'Privileged Access Monitoring (A.8.2)'
    elif 'access_restrictions' in filename_lower or 'wb5' in filename_lower:
        return 'WB5', 'Access Restriction Compliance (A.8.3)'
    elif 'authentication_pam_dashboard' in filename_lower or 'dashboard' in filename_lower or 'wb6' in filename_lower:
        return 'WB6', 'Authentication & PAM Consolidated Dashboard'
    else:
        return 'Unknown', 'Generic Excel Workbook'


# ============================================================================
# EXPECTED SHEET DEFINITIONS
# ============================================================================

EXPECTED_SHEETS = {
    'WB1': [
        "Instructions_Legend", "System_Auth_Inventory", "Auth_Protocol_Analysis",
        "SSO_Integration_Status", "Password_Policy_Compliance", "MFA_Availability_Matrix",
        "Legacy_Auth_Deprecation", "Gap_Analysis", "Evidence_Register", "Approval_Sign_Off"
    ],
    'WB2': [
        "Instructions_Legend", "User_MFA_Enrollment", "MFA_Coverage_By_Type",
        "MFA_Method_Analysis", "Enrollment_Timeline", "MFA_Gaps_Priority",
        "Enrollment_Campaign", "Backup_Method_Status", "Evidence_Register", "Approval_Sign_Off"
    ],
    'WB3': [
        "Instructions_Legend", "Privileged_Account_Inventory", "Admin_Tiering_Matrix",
        "Privileged_User_Roster", "PAM_Vault_Coverage", "MFA_Hardware_Tokens",
        "Credential_Rotation_Status", "Access_Review_Results", "Tier_Isolation_Compliance",
        "Evidence_Register", "Approval_Sign_Off"
    ],
    'WB4': [
        "Instructions_Legend", "Session_Recording_Coverage", "Privileged_Command_Logging",
        "Anomaly_Detection_Rules", "Privileged_Access_Activity", "Alert_Response_Tracking",
        "Off_Hours_Access_Log", "Failed_Login_Analysis", "Tier_Violation_Incidents",
        "Quarterly_Review_Status", "Evidence_Register", "Approval_Sign_Off"
    ],
    'WB5': [
        "Instructions_Legend", "File_System_Permissions", "Database_Access_Controls",
        "Application_RBAC", "API_Access_Controls", "Cloud_IAM_Policies",
        "Encryption_Status", "Network_Segmentation", "Penetration_Test_Results",
        "Gap_Analysis", "Evidence_Register", "Approval_Sign_Off"
    ],
    'WB6': [
        "Executive_Dashboard", "Control_Compliance_Summary", "Critical_Gaps",
        "Risk_Register", "Remediation_Roadmap", "KPI_Metrics",
        "Evidence_Summary", "Action_Items", "Approval_Sign_Off"
    ],
}


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""
    
    workbook_id, workbook_name = detect_workbook_type(filename)
    
    print("=" * 80)
    print(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {filename}")
    print(f"Detected Type: {workbook_id} - {workbook_name}")
    print("=" * 80)
    
    issues_found = []
    warnings_found = []
    
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"\n✓ Workbook loaded successfully")
        print(f"  Sheets found: {len(wb.sheetnames)}")
        print(f"  Sheet names: {', '.join(wb.sheetnames)}")
        
    except Exception as e:
        print(f"\n✗ CRITICAL: Cannot load workbook: {e}")
        return
    
    # ========================================================================
    # CHECK 1: SHEET STRUCTURE VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 1: SHEET STRUCTURE VALIDATION")
    print("=" * 80)
    
    if workbook_id in EXPECTED_SHEETS:
        expected = EXPECTED_SHEETS[workbook_id]
        actual = wb.sheetnames
        
        # Check for missing sheets (flexible matching)
        missing = []
        for exp_sheet in expected:
            found = False
            for act_sheet in actual:
                if exp_sheet.lower().replace("_", "") in act_sheet.lower().replace("_", "").replace(" ", ""):
                    found = True
                    break
            if not found:
                missing.append(exp_sheet)
        
        if missing:
            for sheet in missing:
                warnings_found.append(f"  ⚠ Missing expected sheet: {sheet}")
            print(f"  ⚠ {len(missing)} expected sheets not found (may be renamed)")
        else:
            print(f"  ✓ All expected sheets present for {workbook_id}")
        
        print(f"  Total sheets: {len(actual)} (expected {len(expected)})")
    else:
        print("  ℹ Unknown workbook type - skipping sheet validation")
    
    # ========================================================================
    # CHECK 2: FORMULA VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 2: FORMULA VALIDATION")
    print("=" * 80)
    
    formula_issues = 0
    inter_sheet_refs = {}
    external_refs = set()
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # Check for external workbook references
                    ext_refs = re.findall(r'\[([^\]]+\.xlsx)\]', formula)
                    for ref in ext_refs:
                        external_refs.add(ref)
                    
                    # Check for sheet references
                    if "'" in formula or "!" in formula:
                        sheet_refs = re.findall(r"'?([^'!]+)'?!", formula)
                        for ref in sheet_refs:
                            ref_clean = ref.strip("'")
                            if "[" not in ref_clean and ref_clean not in wb.sheetnames:
                                # Check for close matches
                                close_match = False
                                for actual in wb.sheetnames:
                                    if ref_clean.replace(" ", "_") == actual or ref_clean.replace("_", " ") == actual:
                                        close_match = True
                                        break
                                if not close_match:
                                    issues_found.append(
                                        f"  ✗ {sheet_name}!{cell.coordinate}: "
                                        f"Invalid sheet reference '{ref_clean}'"
                                    )
                                    formula_issues += 1
                    
                    # Track inter-sheet references for report
                    if sheet_name not in inter_sheet_refs:
                        inter_sheet_refs[sheet_name] = set()
                    refs = re.findall(r"'?([^'!\[]+)'?!", formula)
                    for ref in refs:
                        if ref.strip("'") in wb.sheetnames:
                            inter_sheet_refs[sheet_name].add(ref.strip("'"))
    
    if formula_issues == 0:
        print("  ✓ All formulas appear valid")
        if external_refs:
            print(f"  ℹ Found {len(external_refs)} external workbook reference(s):")
            for ref in sorted(external_refs):
                print(f"    - {ref}")
    else:
        print(f"  ✗ Found {formula_issues} formula issue(s)")
    
    if inter_sheet_refs:
        print(f"  ℹ Inter-sheet formula references found in {len(inter_sheet_refs)} sheet(s)")
    
    # ========================================================================
    # CHECK 3: DATA VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 3: DATA VALIDATION")
    print("=" * 80)
    
    validation_issues = 0
    validation_count = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, 'data_validations') and ws.data_validations:
            validations = ws.data_validations.dataValidation
            validation_count += len(validations)
            
            # Check for overlapping validations
            ranges = []
            for dv in validations:
                if hasattr(dv, 'sqref') and dv.sqref:
                    ranges.append((sheet_name, str(dv.sqref)))
    
    if validation_count > 0:
        print(f"  ✓ Found {validation_count} data validation rule(s)")
    else:
        print("  ℹ No data validations found")
    
    # ========================================================================
    # CHECK 4: MERGED CELLS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 4: MERGED CELLS")
    print("=" * 80)
    
    merge_issues = 0
    merge_count = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        merged = ws.merged_cells.ranges if hasattr(ws, 'merged_cells') else []
        merge_count += len(merged)
        
        # Check if merged cells have content in non-top-left cells
        for merged_range in merged:
            min_col, min_row, max_col, max_row = merged_range.bounds
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    if row != min_row or col != min_col:
                        cell = ws.cell(row=row, column=col)
                        if cell.value is not None:
                            warnings_found.append(
                                f"  ⚠ {sheet_name}!{cell.coordinate}: "
                                f"Content in non-top-left cell of merged range"
                            )
                            merge_issues += 1
    
    if merge_count > 0:
        print(f"  ✓ Found {merge_count} merged cell range(s)")
        if merge_issues > 0:
            print(f"  ⚠ {merge_issues} merged cell issue(s) detected")
    else:
        print("  ℹ No merged cells found")
    
    # ========================================================================
    # CHECK 5: NAMED RANGES
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 5: NAMED RANGES")
    print("=" * 80)
    
    if hasattr(wb, 'defined_names') and wb.defined_names:
        named_ranges = list(wb.defined_names.definedName)
        print(f"  ✓ Found {len(named_ranges)} named range(s)")
    else:
        print("  ℹ No named ranges found")
    
    # ========================================================================
    # CHECK 6: WORKBOOK-SPECIFIC VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 6: WORKBOOK-SPECIFIC VALIDATION")
    print("=" * 80)
    
    workbook_specific_check(wb, workbook_id, issues_found, warnings_found)
    
    # ========================================================================
    # DIAGNOSTIC SUMMARY
    # ========================================================================
    print("\n" + "=" * 80)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 80)
    
    if issues_found:
        print(f"\n❌ CRITICAL ISSUES FOUND: {len(issues_found)}")
        for issue in issues_found:
            print(issue)
    
    if warnings_found:
        print(f"\n⚠️ WARNINGS: {len(warnings_found)}")
        for warning in warnings_found:
            print(warning)
    
    if not issues_found and not warnings_found:
        print("\n✅ NO ISSUES DETECTED")
        print("\nThe workbook appears structurally sound.")
        print("If Excel still shows repair warnings, consider:")
        print("  • Excel version compatibility (try Excel 2019+)")
        print("  • Antivirus/security software interference")
        print("  • Network drive or OneDrive sync issues")
        print("  • Excel's overly cautious validation")
        if workbook_id == 'WB6':
            print("  • Unresolved external workbook links (expected before setup)")
    else:
        print_recommendations(issues_found, warnings_found, workbook_id, 
                            formula_issues, validation_issues, merge_issues)
    
    # ========================================================================
    # WORKBOOK-SPECIFIC NOTES
    # ========================================================================
    print_workbook_notes(workbook_id, external_refs)
    
    wb.close()


def workbook_specific_check(wb, workbook_id, issues_found, warnings_found):
    """Perform workbook-specific validation checks."""
    
    if workbook_id == 'WB1':
        # Check System_Auth_Inventory sheet
        if 'System_Auth_Inventory' in wb.sheetnames:
            ws = wb['System_Auth_Inventory']
            # Check for reasonable row count
            if ws.max_row < 154:  # 4 header + 150 data rows
                warnings_found.append(f"  ⚠ System_Auth_Inventory: Expected 150 data rows, found {ws.max_row - 4}")
            else:
                print("  ✓ System_Auth_Inventory has expected row count")
    
    elif workbook_id == 'WB2':
        # Check User_MFA_Enrollment sheet
        if 'User_MFA_Enrollment' in wb.sheetnames:
            ws = wb['User_MFA_Enrollment']
            if ws.max_row < 204:  # 4 header + 200 data rows
                warnings_found.append(f"  ⚠ User_MFA_Enrollment: Expected 200 data rows, found {ws.max_row - 4}")
            else:
                print("  ✓ User_MFA_Enrollment has expected row count")
        
        # Check for MFA coverage formulas
        if 'MFA_Coverage_By_Type' in wb.sheetnames:
            ws = wb['MFA_Coverage_By_Type']
            print("  ✓ MFA_Coverage_By_Type sheet present (formulas will link to User_MFA_Enrollment)")
    
    elif workbook_id == 'WB3':
        # Check Privileged_Account_Inventory
        if 'Privileged_Account_Inventory' in wb.sheetnames:
            ws = wb['Privileged_Account_Inventory']
            if ws.max_row < 154:
                warnings_found.append(f"  ⚠ Privileged_Account_Inventory: Expected 150 data rows, found {ws.max_row - 4}")
            else:
                print("  ✓ Privileged_Account_Inventory has expected row count")
        
        # Check for Admin Tiering Matrix
        if 'Admin_Tiering_Matrix' in wb.sheetnames:
            print("  ✓ Admin_Tiering_Matrix present (Tier 0/1/2 classification)")
    
    elif workbook_id == 'WB4':
        # Check Session_Recording_Coverage
        if 'Session_Recording_Coverage' in wb.sheetnames:
            ws = wb['Session_Recording_Coverage']
            if ws.max_row < 154:
                warnings_found.append(f"  ⚠ Session_Recording_Coverage: Expected 150 data rows, found {ws.max_row - 4}")
            else:
                print("  ✓ Session_Recording_Coverage has expected row count")
    
    elif workbook_id == 'WB5':
        # Check File_System_Permissions
        if 'File_System_Permissions' in wb.sheetnames:
            ws = wb['File_System_Permissions']
            if ws.max_row < 154:
                warnings_found.append(f"  ⚠ File_System_Permissions: Expected 150 data rows, found {ws.max_row - 4}")
            else:
                print("  ✓ File_System_Permissions has expected row count")
    
    elif workbook_id == 'WB6':
        # Dashboard - check for external references
        if 'Executive_Dashboard' in wb.sheetnames:
            print("  ✓ Executive_Dashboard present (consolidation workbook)")
            print("  ℹ This workbook should have external references to WB1-WB5")
    
    else:
        print("  ℹ No workbook-specific validation rules")


def print_recommendations(issues, warnings, workbook_id, formula_issues, 
                         validation_issues, merge_issues):
    """Print recommended fixes based on issues found."""
    print("\n" + "=" * 80)
    print("RECOMMENDED ACTIONS")
    print("=" * 80)
    
    if formula_issues > 0:
        print("\n1. FORMULA FIXES:")
        print("   • Review formulas referencing other sheets")
        print("   • Ensure sheet names match exactly (case-sensitive)")
        print("   • Check for balanced quotes and parentheses")
        if workbook_id == 'WB6':
            print("   • Verify external workbook references are correct")
            print("   • Ensure source workbooks (WB1-WB5) are in same folder")
    
    if validation_issues > 0:
        print("\n2. DATA VALIDATION FIXES:")
        print("   • Remove overlapping data validation ranges")
        print("   • Apply validations to specific ranges, not entire columns")
    
    if merge_issues > 0:
        print("\n3. MERGED CELL FIXES:")
        print("   • Ensure only top-left cell of merged range has content")
        print("   • Clear content from other cells in merged range")
    
    if workbook_id == 'WB6':
        print("\n4. DASHBOARD-SPECIFIC SETUP (WB6):")
        print("   • This workbook consolidates data from WB1-WB5")
        print("   • Ensure source workbooks are in the same folder:")
        print("     - Authentication_Inventory_Assessment_NORMALIZED.xlsx (WB1)")
        print("     - MFA_Coverage_Assessment_NORMALIZED.xlsx (WB2)")
        print("     - Privileged_Accounts_Assessment_NORMALIZED.xlsx (WB3)")
        print("     - Privileged_Monitoring_Assessment_NORMALIZED.xlsx (WB4)")
        print("     - Access_Restrictions_Assessment_NORMALIZED.xlsx (WB5)")
        print("   • Run normalize_assessment_files_a8235.py to create normalized versions")


def print_workbook_notes(workbook_id, external_refs):
    """Print workbook-specific notes and expected structure."""
    print("\n" + "=" * 80)
    print(f"WORKBOOK-SPECIFIC NOTES: {workbook_id}")
    print("=" * 80)
    
    if workbook_id == 'WB1':
        print("\nAuthentication Inventory & Methods Assessment (A.8.5):")
        print("  • 10 sheets for authentication mechanism tracking")
        print("  • 150 system rows in System_Auth_Inventory")
        print("  • Authentication protocol analysis (SAML, OAuth, OIDC, Kerberos)")
        print("  • SSO integration status (target 90%+ adoption)")
        print("  • Password policy compliance verification")
        print("  • MFA availability matrix per system")
        print("  • Legacy authentication deprecation tracking")
        print("  • Evidence Register (30 entries)")
        print("  • 3-level approval workflow")
    
    elif workbook_id == 'WB2':
        print("\nMFA Coverage Assessment (A.8.5):")
        print("  • 10 sheets for MFA enrollment tracking")
        print("  • 200 user rows in User_MFA_Enrollment")
        print("  • Coverage metrics by user type (privileged/remote/standard)")
        print("  • MFA method security analysis (hardware token > app > SMS)")
        print("  • 12-month enrollment timeline tracking")
        print("  • Gap prioritization (privileged users without MFA = critical)")
        print("  • Enrollment campaign tracking (30 rows)")
        print("  • Backup method status verification (50 rows)")
        print("  • NIS2 Article 21(2)(e) compliance emphasis (MFA MANDATORY)")
        print("  • Evidence Register (30 entries)")
    
    elif workbook_id == 'WB3':
        print("\nPrivileged Accounts Inventory (A.8.2):")
        print("  • 11 sheets for privileged account management")
        print("  • 150 account rows in Privileged_Account_Inventory")
        print("  • Admin Tiering Matrix (Tier 0/1/2 classification) ⭐")
        print("  • Privileged user roster (100 users)")
        print("  • PAM vault coverage tracking (target 100% shared accounts)")
        print("  • Hardware MFA token deployment for Tier 0 (40 rows)")
        print("  • Credential rotation status (100 rows)")
        print("  • Quarterly access review results (80 rows)")
        print("  • Tier isolation violation tracking (target: ZERO)")
        print("  • Evidence Register (30 entries)")
    
    elif workbook_id == 'WB4':
        print("\nPrivileged Access Monitoring (A.8.2):")
        print("  • 12 sheets for monitoring and anomaly detection")
        print("  • 150 rows session recording coverage tracking")
        print("  • Command logging verification (100 rows)")
        print("  • Anomaly detection rules (30 alert definitions)")
        print("  • Privileged access activity log (200 entries)")
        print("  • Alert response tracking (100 alerts, <30 min SLA critical)")
        print("  • Off-hours access log (80 entries)")
        print("  • Failed login analysis (100 attempts)")
        print("  • Tier violation incidents (50 rows, should be ZERO)")
        print("  • Quarterly review status (40 periods)")
        print("  • Evidence Register (30 entries)")
    
    elif workbook_id == 'WB5':
        print("\nAccess Restriction Compliance (A.8.3):")
        print("  • 12 sheets for technical access control assessment")
        print("  • 150 rows file system permission audits")
        print("  • Database access control verification (100 rows)")
        print("  • Application RBAC assessment (80 rows)")
        print("  • API access control evaluation (60 rows)")
        print("  • Cloud IAM policy review (80 rows)")
        print("  • Encryption status tracking (100 rows)")
        print("  • Network segmentation compliance (50 rows)")
        print("  • Penetration test results (50 findings)")
        print("  • Gap analysis (80 entries)")
        print("  • Evidence Register (30 entries)")
    
    elif workbook_id == 'WB6':
        print("\nAuthentication & PAM Consolidated Dashboard:")
        print("  • 9 sheets - MASTER CONSOLIDATION WORKBOOK")
        print("  • Executive Dashboard with compliance metrics")
        print("  • Control-by-control summary (A.8.2, A.8.3, A.8.5)")
        print("  • Critical gaps prioritization (30 items)")
        print("  • Risk register integration (50 entries)")
        print("  • Remediation roadmap (40 items)")
        print("  • KPI metrics tracking")
        print("  • Evidence summary (100 items across all workbooks)")
        print("  • Action items (50 entries)")
        print("  • CISO approval workflow")
        print("\n  ⚠ This dashboard may show #REF errors until:")
        print("    1. All 5 source workbooks (WB1-WB5) are completed")
        print("    2. Files are normalized using normalize_assessment_files_a8235.py")
        print("    3. Normalized files are in same folder as dashboard")
        print("    4. Links are updated in Excel (Edit Links → Update Values)")
    
    else:
        print("\nGeneric Excel Workbook")
        print("  • No specific validation rules defined for this type")
        print("  • Standard Excel health checks applied")
    
    print("\n" + "=" * 80)


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("=" * 70)
        print("ISMS Controls A.8.2/3/5 Authentication & PAM - Excel Sanity Checker")
        print("=" * 70)
        print("\nUsage: python3 excel_sanity_check_a8235.py <filename.xlsx>")
        print("\nExamples:")
        print("  python3 excel_sanity_check_a8235.py Authentication_Inventory_Assessment_20250111.xlsx")
        print("  python3 excel_sanity_check_a8235.py MFA_Coverage_Assessment_20250111.xlsx")
        print("  python3 excel_sanity_check_a8235.py Privileged_Accounts_Assessment_20250111.xlsx")
        print("  python3 excel_sanity_check_a8235.py Privileged_Monitoring_Assessment_20250111.xlsx")
        print("  python3 excel_sanity_check_a8235.py Access_Restrictions_Assessment_20250111.xlsx")
        print("  python3 excel_sanity_check_a8235.py Authentication_PAM_Dashboard_20250111.xlsx")
        print("\nSupported workbook types:")
        print("  WB1 - Authentication Inventory & Methods (A.8.5)")
        print("  WB2 - MFA Coverage Assessment (A.8.5)")
        print("  WB3 - Privileged Accounts Inventory (A.8.2)")
        print("  WB4 - Privileged Access Monitoring (A.8.2)")
        print("  WB5 - Access Restriction Compliance (A.8.3)")
        print("  WB6 - Authentication & PAM Consolidated Dashboard")
        print("\n'Evidence > Theater' - Systems Engineering ISMS")
        sys.exit(1)
    
    filename = sys.argv[1]
    
    try:
        check_workbook_health(filename)
    except FileNotFoundError:
        print(f"\n❌ ERROR: File not found: {filename}")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()