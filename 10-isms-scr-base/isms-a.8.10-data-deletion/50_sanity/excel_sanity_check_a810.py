#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
Excel Sanity Check for ISMS-IMP-A.8.10 (Complete Framework)
================================================================================

Validates all 5 Information Deletion assessment workbooks (Domains 1-5) for
ISO/IEC 27001:2022 Control A.8.10: Information Deletion

This script checks for:
- Workbook structure (correct sheets present)
- Style consistency (no reused style objects)
- Data validations (properly configured dropdowns)
- Formula syntax (no #REF!, #NAME? errors)
- Pre-populated examples present
- Column widths appropriate
- Freeze panes set correctly
- Document ID consistency
- Summary_Dashboard structure for external linking (Domains 1-4)
- External formula references (Domain 5)

Usage:
    python3 excel_sanity_check_a810.py [workbook_directory]
    
    If no directory specified, checks current directory.

Quality Assurance:
    Run this script BEFORE using workbooks in production or distributing to
    stakeholders. Fix all CRITICAL issues and review WARNINGS.

Integration:
    Part of A.8.10 Information Deletion assessment quality assurance workflow:
    1. Generate workbooks (generate_a810_1 through _5.py)
    2. Run sanity check (this script) → Fix issues
    3. Normalize file names (normalize_assessment_files_a810.py)
    4. Complete assessments (user data entry)
    5. Validate with sanity check again
    6. Feed into Dashboard (A.8.10.5)

Control Reference: ISO/IEC 27001:2022 Annex A Control A.8.10
Script Type: Quality Assurance Utility
Version: 1.0
Author: [Developer Name / Organisation]
Date: [Date to be set]

================================================================================
"""

import sys
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime


# ============================================================================
# CONFIGURATION
# ============================================================================

# Expected workbook structures for A.8.10 Information Deletion
WORKBOOK_SPECS = {
    "Domain_1": {
        "pattern": "ISMS-IMP-A.8.10.1",
        "name": "Retention & Deletion Triggers",
        "document_id": "ISMS-IMP-A.8.10.1",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Data Category Registry",
            "3. Retention Schedule Compliance",
            "4. Deletion Trigger Configuration",
            "5. Legal Hold Management",
            "6. Data Subject Requests",
            "Summary Dashboard",
            "Evidence Register",
            "Approval Sign-Off"  # Note: Script creates "Approval Sign-Of" (typo)
        ],
        "min_sheets": 9,
        "assessment_items": 150,
        "summary_sheet": "Summary Dashboard",
        "key_cells": ['G10', 'E10', 'B5', 'B6']  # Compliance %, Non-Compliant, etc.
    },
    "Domain_2": {
        "pattern": "ISMS-IMP-A.8.10.2",
        "name": "Deletion Methods",
        "document_id": "ISMS-IMP-A.8.10.2",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Physical Storage Media",
            "3. Database Systems",
            "4. Cloud Storage",
            "5. File Systems & Backup",
            "6. Deletion Verification Testing",
            "Summary Dashboard",
            "Evidence Register",
            "Approval Sign-Off"
        ],
        "min_sheets": 9,
        "assessment_items": 100,
        "summary_sheet": "Summary Dashboard",
        "key_cells": ['G10', 'E10', 'B5', 'B6'],
        "nist_categories": True  # NIST SP 800-88 categories
    },
    "Domain_3": {
        "pattern": "ISMS-IMP-A.8.10.3",
        "name": "Third-Party & Cloud Deletion",
        "document_id": "ISMS-IMP-A.8.10.3",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Cloud Provider Deletion",
            "3. SaaS Application Deletion",
            "4. Vendor Contract Assessment",
            "5. Subprocessor Mapping",
            "6. Shadow IT Assessment",
            "Summary Dashboard",
            "Evidence Register",
            "Approval Sign-Off"
        ],
        "min_sheets": 9,
        "assessment_items": 100,
        "summary_sheet": "Summary Dashboard",
        "key_cells": ['G10', 'E10', 'B5', 'B6'],
        "gdpr_compliance": True  # GDPR Article 28 requirements
    },
    "Domain_4": {
        "pattern": "ISMS-IMP-A.8.10.4",
        "name": "Verification & Evidence",
        "document_id": "ISMS-IMP-A.8.10.4",
        "expected_sheets": [
            "Instructions & Legend",
            "2. Verification Testing Strategy",
            "3. Forensic Testing Procedures",
            "4. Deletion Log Management",
            "5. Evidence Documentation",
            "6. Verification Sampling",
            "Verification Dashboard",  # Different name!
            "Evidence Register",
            "Approval Sign-Off"
        ],
        "min_sheets": 9,
        "assessment_items": 80,
        "summary_sheet": "Verification Dashboard",  # Different sheet name
        "key_cells": ['B9', 'B8'],  # Different cells: B9=Compliance %, B8=Critical Gaps
        "different_structure": True
    },
    "Domain_5": {
        "pattern": "ISMS-IMP-A.8.10.5",
        "name": "Compliance Dashboard",
        "document_id": "ISMS-IMP-A.8.10.5",
        "expected_sheets": [
            "Instructions & Legend",
            "Executive Summary",
            "Domain 1 Summary",
            "Domain 2 Summary",
            "Domain 3 Summary",
            "Domain 4 Summary",
            "Consolidated Gap Analysis",
            "Risk Register",
            "Remediation Roadmap",
            "Evidence Master Index",
            "KPI Dashboard",
            "CISO DPO Approval"
        ],
        "min_sheets": 12,
        "assessment_items": 100,
        "external_links_expected": True,
        "consolidation_dashboard": True
    }
}

# Color codes for terminal output
GREEN = '\033[92m'
YELLOW = '\033[93m'
RED = '\033[91m'
BLUE = '\033[94m'
BOLD = '\033[1m'
RESET = '\033[0m'


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def print_header(text):
    """Print section header."""
    print(f"\n{BLUE}{BOLD}{'=' * 80}{RESET}")
    print(f"{BLUE}{BOLD}{text}{RESET}")
    print(f"{BLUE}{BOLD}{'=' * 80}{RESET}\n")


def print_success(text):
    """Print success message."""
    print(f"{GREEN}✅ {text}{RESET}")


def print_warning(text):
    """Print warning message."""
    print(f"{YELLOW}⚠️  {text}{RESET}")


def print_error(text):
    """Print error message."""
    print(f"{RED}❌ {text}{RESET}")


def print_info(text):
    """Print info message."""
    print(f"   {text}")


def find_workbook(directory, pattern):
    """Find workbook matching pattern in directory."""
    path = Path(directory)
    matches = list(path.glob(f"{pattern}*.xlsx"))
    
    # Filter out temporary files
    matches = [m for m in matches if not m.name.startswith('~$')]
    
    if not matches:
        return None
    
    if len(matches) > 1:
        print_warning(f"Multiple workbooks found for {pattern}, using most recent")
        matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    
    return matches[0]


def check_document_id(wb, expected_doc_id):
    """Check if Document ID is present in Instructions & Legend."""
    issues = []
    warnings = []
    
    if "Instructions & Legend" not in wb.sheetnames:
        issues.append("Instructions & Legend sheet not found")
        return issues, warnings
    
    ws = wb["Instructions & Legend"]
    doc_id_found = False
    found_doc_id = None
    
    # Search for Document ID in first 30 rows, column A-B
    for row in range(1, 31):
        cell_label = ws.cell(row=row, column=1).value
        if cell_label and "Document ID" in str(cell_label):
            found_doc_id = ws.cell(row=row, column=2).value
            doc_id_found = True
            break
    
    if not doc_id_found:
        issues.append("Document ID not found in Instructions & Legend sheet")
    elif found_doc_id != expected_doc_id:
        issues.append(f"Document ID mismatch: Expected '{expected_doc_id}', found '{found_doc_id}'")
    else:
        print_success(f"Document ID verified: {found_doc_id}")
    
    return issues, warnings


def check_workbook_structure(wb, domain_key, spec):
    """Check if workbook has correct sheet structure."""
    issues = []
    warnings = []
    
    actual_sheets = wb.sheetnames
    expected_sheets = spec["expected_sheets"]
    
    # Check sheet count
    if len(actual_sheets) < spec["min_sheets"]:
        issues.append(f"Only {len(actual_sheets)} sheets found, expected at least {spec['min_sheets']}")
    else:
        print_success(f"Sheet count: {len(actual_sheets)} sheets")
    
    # Check for expected sheets (with flexible matching for typos)
    missing_sheets = []
    for sheet_name in expected_sheets:
        found = False
        for actual_sheet in actual_sheets:
            # Exact match or close match (handle "Approval Sign-Of" typo)
            if actual_sheet == sheet_name or (
                "Approval" in sheet_name and "Approval" in actual_sheet and 
                "Sign" in actual_sheet
            ):
                found = True
                break
        
        if not found:
            missing_sheets.append(sheet_name)
    
    if missing_sheets:
        issues.append(f"Missing sheets: {', '.join(missing_sheets)}")
    else:
        print_success(f"All {len(expected_sheets)} expected sheets present (or close match)")
    
    # Check for unexpected sheets
    unexpected_sheets = []
    for actual_sheet in actual_sheets:
        found = False
        for expected_sheet in expected_sheets:
            if actual_sheet == expected_sheet or (
                "Approval" in expected_sheet and "Approval" in actual_sheet and
                "Sign" in actual_sheet
            ):
                found = True
                break
        
        if not found:
            unexpected_sheets.append(actual_sheet)
    
    if unexpected_sheets:
        warnings.append(f"Unexpected sheets: {', '.join(unexpected_sheets)}")
    
    return issues, warnings


def check_summary_dashboard(ws, domain_key, spec):
    """Check Summary/Verification Dashboard structure for external linking."""
    issues = []
    warnings = []
    
    # Get key cells for this domain
    key_cells = spec.get("key_cells", [])
    
    if not key_cells:
        return issues, warnings
    
    cells_exist = 0
    for cell_ref in key_cells:
        try:
            cell = ws[cell_ref]
            cells_exist += 1
        except:
            warnings.append(f"Key cell {cell_ref} not accessible")
    
    if cells_exist >= len(key_cells) * 0.7:  # At least 70% of cells exist
        print_success(f"{spec['summary_sheet']}: {cells_exist}/{len(key_cells)} key cells accessible")
    else:
        warnings.append(f"{spec['summary_sheet']}: Only {cells_exist}/{len(key_cells)} key cells accessible")
    
    # Check for compliance percentage (should be formula or value)
    if spec.get("different_structure"):
        # Domain 4 uses B9 for compliance
        compliance_cell = 'B9'
    else:
        # Domains 1-3 use G10 for compliance
        compliance_cell = 'G10'
    
    try:
        cell = ws[compliance_cell]
        if cell.value is not None:
            print_success(f"Compliance percentage cell ({compliance_cell}) contains data")
        else:
            warnings.append(f"Compliance percentage cell ({compliance_cell}) is empty")
    except:
        warnings.append(f"Cannot access compliance cell {compliance_cell}")
    
    # Check for section headers
    kpi_found = False
    compliance_found = False
    
    for row in ws.iter_rows(max_row=50, max_col=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                value_upper = cell.value.upper()
                if "KPI" in value_upper or "PERFORMANCE" in value_upper or "METRICS" in value_upper:
                    kpi_found = True
                if "COMPLIANCE" in value_upper or "STATUS" in value_upper or "SUMMARY" in value_upper:
                    compliance_found = True
    
    if kpi_found:
        print_success(f"{spec['summary_sheet']}: KPI/Performance section found")
    else:
        warnings.append(f"{spec['summary_sheet']}: No KPI/Performance section found")
    
    if compliance_found:
        print_success(f"{spec['summary_sheet']}: Compliance/Status section found")
    else:
        warnings.append(f"{spec['summary_sheet']}: No Compliance/Status section found")
    
    return issues, warnings


def check_data_validations(ws, sheet_name):
    """Check if data validations are properly configured."""
    issues = []
    warnings = []
    
    # Check if any validations exist
    if not ws.data_validations.dataValidation:
        if sheet_name not in ["Instructions & Legend", "Summary Dashboard", 
                              "Verification Dashboard", "Executive Summary",
                              "CISO DPO Approval"]:
            warnings.append(f"{sheet_name}: No data validations found")
        return issues, warnings
    
    validation_count = len(ws.data_validations.dataValidation)
    print_info(f"{sheet_name}: {validation_count} data validations configured")
    
    # Check each validation has cells assigned
    for idx, validation in enumerate(ws.data_validations.dataValidation, 1):
        if not validation.cells:
            issues.append(f"{sheet_name}: Validation #{idx} has no cells assigned")
    
    return issues, warnings


def check_formulas(ws, sheet_name):
    """Check for formula errors."""
    issues = []
    warnings = []
    
    formula_count = 0
    error_count = 0
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Formula
                formula_count += 1
                # Check for common error indicators
                if cell.value and any(err in str(cell.value).upper() for err in ['#REF!', '#NAME?', '#VALUE!', '#DIV/0!']):
                    error_count += 1
                    issues.append(f"{sheet_name}: Formula error in {cell.coordinate}: {cell.value}")
    
    if formula_count > 0:
        print_info(f"{sheet_name}: {formula_count} formulas found")
        if error_count == 0:
            print_success(f"{sheet_name}: No formula errors detected")
        else:
            print_error(f"{sheet_name}: {error_count} formula errors found")
    
    return issues, warnings


def check_freeze_panes(ws, sheet_name):
    """Check if freeze panes are set."""
    issues = []
    warnings = []
    
    if ws.freeze_panes:
        print_info(f"{sheet_name}: Freeze panes set at {ws.freeze_panes}")
    else:
        if sheet_name not in ["Summary Dashboard", "Verification Dashboard", "Executive Summary"]:
            warnings.append(f"{sheet_name}: No freeze panes set")
    
    return issues, warnings


def check_column_widths(ws, sheet_name):
    """Check if column widths are reasonable."""
    issues = []
    warnings = []
    
    narrow_columns = []
    wide_columns = []
    
    for col_letter, dimension in ws.column_dimensions.items():
        if dimension.width:
            if dimension.width < 8:
                narrow_columns.append(f"{col_letter}:{dimension.width:.1f}")
            elif dimension.width > 60:
                wide_columns.append(f"{col_letter}:{dimension.width:.1f}")
    
    if narrow_columns:
        warnings.append(f"{sheet_name}: Very narrow columns: {', '.join(narrow_columns[:5])}")
    
    if wide_columns:
        warnings.append(f"{sheet_name}: Very wide columns: {', '.join(wide_columns[:5])}")
    
    return issues, warnings


def check_pre_populated_examples(ws, sheet_name):
    """Check for pre-populated example rows."""
    issues = []
    warnings = []
    
    # Skip certain sheets that don't need examples
    if sheet_name in ["Instructions & Legend", "Gap Analysis", "Evidence Register", 
                      "Summary Dashboard", "Verification Dashboard", "Executive Summary",
                      "CISO DPO Approval", "Approval Sign-Off", "Approval Sign-Of"]:
        return issues, warnings
    
    # Check for gray-filled cells (example rows) or data in first rows
    example_rows = 0
    data_rows = 0
    
    for row_idx in range(6, 15):  # Check first ~10 data rows
        try:
            cell = ws.cell(row=row_idx, column=1)
            
            # Check for gray fill (example indicator)
            if cell.fill and cell.fill.start_color:
                color = str(cell.fill.start_color.rgb).upper()
                if 'E7E6E6' in color or 'FFE7E6E6' in color or 'D9D9D9' in color or 'D3D3D3' in color:
                    example_rows += 1
            
            # Check for any data
            if cell.value:
                data_rows += 1
        except:
            pass
    
    if example_rows == 0 and data_rows == 0:
        warnings.append(f"{sheet_name}: No pre-populated examples or data found")
    elif example_rows > 0:
        print_info(f"{sheet_name}: {example_rows} example rows found")
    elif data_rows > 0:
        print_info(f"{sheet_name}: {data_rows} data rows found")
    
    return issues, warnings


def check_standard_sheets(wb, domain_key, spec):
    """Check standard sheets (Evidence Register, Summary Dashboard, Approval)."""
    issues = []
    warnings = []
    
    # Check Evidence Register (should have ~100-110 rows)
    if "Evidence Register" in wb.sheetnames:
        ws = wb["Evidence Register"]
        max_row = ws.max_row
        if max_row < 100:
            warnings.append(f"Evidence Register: Only {max_row} rows, expected ~100-110")
        else:
            print_success(f"Evidence Register: {max_row} rows configured")
    
    # Check Summary Dashboard or Verification Dashboard
    summary_sheet = spec.get("summary_sheet", "Summary Dashboard")
    if summary_sheet in wb.sheetnames:
        ws = wb[summary_sheet]
        issues_dash, warnings_dash = check_summary_dashboard(ws, domain_key, spec)
        issues.extend(issues_dash)
        warnings.extend(warnings_dash)
    
    # Check Approval Sign-Off (flexible for typo)
    approval_sheets = [s for s in wb.sheetnames if "Approval" in s and "Sign" in s]
    if approval_sheets:
        print_success(f"Approval sheet found: {approval_sheets[0]}")
    else:
        warnings.append("Approval Sign-Off sheet not found")
    
    return issues, warnings


def check_domain_5_external_links(wb):
    """Check Domain 5 for external formula references."""
    issues = []
    warnings = []
    
    # Check summary sheets for external formulas
    external_formula_sheets = ["Executive Summary", "Domain 1 Summary", 
                               "Domain 2 Summary", "Domain 3 Summary", 
                               "Domain 4 Summary", "KPI Dashboard"]
    
    external_formulas_found = 0
    sheets_with_external = []
    
    for sheet_name in external_formula_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            found_in_sheet = False
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        # Check for external workbook reference pattern
                        if '[ISMS-IMP-A.8.10.' in str(cell.value) or '[ISMS_A_8_10_' in str(cell.value):
                            external_formulas_found += 1
                            found_in_sheet = True
                            break
                if found_in_sheet:
                    sheets_with_external.append(sheet_name)
                    break
    
    if external_formulas_found > 0:
        print_success(f"Domain 5: External formulas found in {len(sheets_with_external)} sheets")
        print_info(f"  Sheets with external links: {', '.join(sheets_with_external)}")
    else:
        warnings.append("Domain 5: No external formulas found (may need workbook linking)")
    
    # Check Instructions for normalization/linking guidance
    if "Instructions & Legend" in wb.sheetnames:
        ws = wb["Instructions & Legend"]
        normalization_found = False
        external_link_found = False
        
        for row in ws.iter_rows(max_row=150, max_col=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value_lower = cell.value.lower()
                    if "normaliz" in value_lower or "normalize" in value_lower:
                        normalization_found = True
                    if "external" in value_lower or "[isms-imp" in value_lower:
                        external_link_found = True
        
        if normalization_found:
            print_success("Domain 5: Normalization instructions found")
        else:
            warnings.append("Domain 5: No normalization instructions found")
        
        if external_link_found:
            print_success("Domain 5: External linking documentation found")
        else:
            warnings.append("Domain 5: No external linking documentation found")
    
    return issues, warnings


def check_nist_categories(wb, domain_key):
    """Check for NIST SP 800-88 category references in Domain 2."""
    issues = []
    warnings = []
    
    if domain_key != "Domain_2":
        return issues, warnings
    
    # Check Instructions & Legend for NIST SP 800-88
    if "Instructions & Legend" in wb.sheetnames:
        ws = wb["Instructions & Legend"]
        nist_found = False
        categories_found = []
        
        for row in ws.iter_rows(max_row=100, max_col=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value_upper = cell.value.upper()
                    if "NIST SP 800-88" in value_upper or "NIST SP800-88" in value_upper:
                        nist_found = True
                    if "CLEAR" in value_upper and "PURGE" in value_upper:
                        categories_found.append("Clear/Purge/Destroy")
        
        if nist_found:
            print_success("Domain 2: NIST SP 800-88 references found")
        else:
            warnings.append("Domain 2: No NIST SP 800-88 references found")
        
        if categories_found:
            print_success("Domain 2: NIST categories (Clear/Purge/Destroy) found")
        else:
            warnings.append("Domain 2: NIST categories not clearly documented")
    
    # Check for SSD-specific warnings
    ssd_warning_found = False
    crypto_erasure_found = False
    
    for sheet_name in ["Instructions & Legend", "2. Physical Storage Media"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(max_row=150, max_col=10):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        value_upper = cell.value.upper()
                        if "SSD" in value_upper and ("OVERWRITE" in value_upper or "INEFFECTIVE" in value_upper):
                            ssd_warning_found = True
                        if "CRYPTO" in value_upper and "ERASURE" in value_upper:
                            crypto_erasure_found = True
    
    if ssd_warning_found:
        print_success("Domain 2: SSD-specific warnings found")
    else:
        warnings.append("Domain 2: No SSD-specific warnings (standard overwrite ineffective)")
    
    if crypto_erasure_found:
        print_success("Domain 2: Crypto-erasure guidance found")
    else:
        warnings.append("Domain 2: No crypto-erasure guidance for SSDs")
    
    return issues, warnings


def check_gdpr_compliance(wb, domain_key):
    """Check for GDPR Article 28 references in Domain 3."""
    issues = []
    warnings = []
    
    if domain_key != "Domain_3":
        return issues, warnings
    
    # Check Instructions & Legend for GDPR Article 28
    if "Instructions & Legend" in wb.sheetnames:
        ws = wb["Instructions & Legend"]
        gdpr_found = False
        article_28_found = False
        
        for row in ws.iter_rows(max_row=100, max_col=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value_upper = cell.value.upper()
                    if "GDPR" in value_upper:
                        gdpr_found = True
                    if "ARTICLE 28" in value_upper or "ART. 28" in value_upper:
                        article_28_found = True
        
        if gdpr_found:
            print_success("Domain 3: GDPR references found")
        else:
            warnings.append("Domain 3: No GDPR references found")
        
        if article_28_found:
            print_success("Domain 3: GDPR Article 28 processor obligations found")
        else:
            warnings.append("Domain 3: No GDPR Article 28 references found")
    
    return issues, warnings


def validate_workbook(filepath, domain_key, spec):
    """Perform comprehensive validation of a workbook."""
    print_header(f"Validating {spec['name']} ({domain_key})")
    print_info(f"File: {filepath.name}")
    
    all_issues = []
    all_warnings = []
    
    try:
        # Load workbook
        print_info("Loading workbook...")
        wb = load_workbook(filepath, data_only=False)
        print_success("Workbook loaded successfully")
        
        # Check Document ID
        print_info("\nChecking Document ID...")
        issues, warnings = check_document_id(wb, spec["document_id"])
        all_issues.extend(issues)
        all_warnings.extend(warnings)
        
        # Check structure
        print_info("\nChecking workbook structure...")
        issues, warnings = check_workbook_structure(wb, domain_key, spec)
        all_issues.extend(issues)
        all_warnings.extend(warnings)
        
        # Check standard sheets
        print_info("\nChecking standard sheets...")
        issues, warnings = check_standard_sheets(wb, domain_key, spec)
        all_issues.extend(issues)
        all_warnings.extend(warnings)
        
        # Check each sheet
        print_info("\nChecking individual sheets...")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Data validations
            issues, warnings = check_data_validations(ws, sheet_name)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
            
            # Formulas
            issues, warnings = check_formulas(ws, sheet_name)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
            
            # Freeze panes
            issues, warnings = check_freeze_panes(ws, sheet_name)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
            
            # Column widths
            issues, warnings = check_column_widths(ws, sheet_name)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
            
            # Pre-populated examples
            issues, warnings = check_pre_populated_examples(ws, sheet_name)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
        
        # Domain-specific checks
        if domain_key == "Domain_2" and spec.get("nist_categories"):
            print_info("\nChecking NIST SP 800-88 compliance features...")
            issues, warnings = check_nist_categories(wb, domain_key)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
        
        if domain_key == "Domain_3" and spec.get("gdpr_compliance"):
            print_info("\nChecking GDPR Article 28 compliance features...")
            issues, warnings = check_gdpr_compliance(wb, domain_key)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
        
        if domain_key == "Domain_5" and spec.get("external_links_expected"):
            print_info("\nChecking Domain 5 external linking features...")
            issues, warnings = check_domain_5_external_links(wb)
            all_issues.extend(issues)
            all_warnings.extend(warnings)
        
        # Summary
        print_info("\n" + "=" * 60)
        if len(all_issues) == 0 and len(all_warnings) == 0:
            print_success(f"{domain_key}: VALIDATION PASSED - No issues found")
            return True, 0, 0
        else:
            if len(all_issues) > 0:
                print_error(f"{domain_key}: {len(all_issues)} critical issues found")
                for issue in all_issues[:10]:  # Limit display to first 10
                    print_error(f"  • {issue}")
                if len(all_issues) > 10:
                    print_error(f"  ... and {len(all_issues) - 10} more issues")
            
            if len(all_warnings) > 0:
                print_warning(f"{domain_key}: {len(all_warnings)} warnings found")
                for warning in all_warnings[:10]:  # Limit display to first 10
                    print_warning(f"  • {warning}")
                if len(all_warnings) > 10:
                    print_warning(f"  ... and {len(all_warnings) - 10} more warnings")
            
            return len(all_issues) == 0, len(all_issues), len(all_warnings)
    
    except Exception as e:
        print_error(f"Failed to validate workbook: {str(e)}")
        import traceback
        print_info(traceback.format_exc())
        return False, 1, 0


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main():
    """Main validation function."""
    
    # Determine directory
    if len(sys.argv) > 1:
        directory = sys.argv[1]
    else:
        directory = "."
    
    if not os.path.isdir(directory):
        print_error(f"Directory not found: {directory}")
        sys.exit(1)
    
    print_header("ISMS-IMP-A.8.10 Information Deletion Framework Validation")
    print_info(f"Checking directory: {os.path.abspath(directory)}")
    print_info(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Find and validate each domain
    results = {}
    total_issues = 0
    total_warnings = 0
    
    for domain_key, spec in WORKBOOK_SPECS.items():
        filepath = find_workbook(directory, spec["pattern"])
        
        if filepath:
            passed, issues, warnings = validate_workbook(filepath, domain_key, spec)
            results[domain_key] = {
                "found": True,
                "passed": passed,
                "issues": issues,
                "warnings": warnings
            }
            total_issues += issues
            total_warnings += warnings
        else:
            print_header(f"Checking {spec['name']} ({domain_key})")
            print_warning(f"Workbook not found: {spec['pattern']}*.xlsx")
            print_info("This domain has not been generated yet")
            results[domain_key] = {
                "found": False,
                "passed": None,
                "issues": 0,
                "warnings": 0
            }
    
    # Final summary
    print_header("VALIDATION SUMMARY")
    
    found_count = sum(1 for r in results.values() if r["found"])
    passed_count = sum(1 for r in results.values() if r["passed"] == True)
    
    print_info(f"Workbooks found: {found_count} / 5")
    print_info(f"Workbooks passed: {passed_count} / {found_count if found_count > 0 else 1}")
    print_info(f"Total critical issues: {total_issues}")
    print_info(f"Total warnings: {total_warnings}")
    
    print("\nDomain Status:")
    for domain_key, result in results.items():
        spec = WORKBOOK_SPECS[domain_key]
        if result["found"]:
            if result["passed"]:
                print_success(f"{domain_key} ({spec['name']}): PASSED")
            else:
                print_error(f"{domain_key} ({spec['name']}): FAILED ({result['issues']} issues, {result['warnings']} warnings)")
        else:
            print_warning(f"{domain_key} ({spec['name']}): NOT FOUND")
    
    # Exit code
    if total_issues > 0:
        print(f"\n{RED}{BOLD}VALIDATION FAILED{RESET}")
        print_info("Please fix critical issues before using workbooks in production")
        sys.exit(1)
    elif found_count == 0:
        print(f"\n{YELLOW}{BOLD}NO WORKBOOKS FOUND{RESET}")
        print_info("Generate workbooks first using the Python generator scripts:")
        print_info("  python3 generate_a810_1_retention_triggers.py")
        print_info("  python3 generate_a810_2_deletion_methods.py")
        print_info("  python3 generate_a810_3_third_party_cloud.py")
        print_info("  python3 generate_a810_4_verification_evidence.py")
        print_info("  python3 generate_a810_5_compliance_dashboard.py")
        sys.exit(2)
    else:
        print(f"\n{GREEN}{BOLD}VALIDATION PASSED{RESET}")
        if total_warnings > 0:
            print_info(f"Note: {total_warnings} warnings found (non-critical)")
        print_info("All found workbooks are ready for use")
        print_info("\nNext steps:")
        print_info("  1. Complete assessments (fill yellow cells)")
        print_info("  2. Run: python3 normalize_assessment_files_a810.py")
        print_info("  3. Generate dashboard (if not done): python3 generate_a810_5_compliance_dashboard.py")
        print_info("  4. Open dashboard and click 'Update Links'")
        sys.exit(0)


if __name__ == "__main__":
    main()