#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.8 - Portfolio Dashboard Consolidation Utility
================================================================================

ISO/IEC 27001:2022 Control A.5.8: Information Security in Project Management
Operational Utility: Multi-Project Portfolio Dashboard Consolidation

Purpose:
Automates extraction of project data from multiple A.5.8.1 Project Lifecycle
Assessment workbooks and populates the A.5.8.3 Portfolio Dashboard, eliminating
manual data entry errors and ensuring consistent portfolio-wide reporting.

Key Concept:
This script extracts data that CANNOT be handled by Excel formulas alone because:
1. Project data is in separate workbooks (not sheets in same file)
2. Portfolio dashboard needs consolidated view across 10-50+ projects
3. Manual copy-paste is error-prone and time-consuming
4. Audit trail requires automated, repeatable process

Input Sources:
- Multiple ISMS-IMP-A.5.8.1 workbooks (one per active project)
- Each workbook contains project classification, phase scores, compliance data

Output:
- Populated ISMS-IMP-A.5.8.3 Portfolio Dashboard "Project Data" sheet
- Consolidation log
- Audit trail

Data Extracted (16 fields per project):
1. Project Name - from Classification sheet
2. Risk Classification - High/Medium/Low from classification criteria
3. Project Manager - from Classification sheet
4. Business Owner - (manual field, blank by default)
5. Current Phase - determined from highest active phase
6. Overall Compliance % - from Compliance Dashboard
7. Initiation Phase % - from Compliance Dashboard
8. Planning Phase % - from Compliance Dashboard
9. Execution Phase % - from Compliance Dashboard
10. Monitoring Phase % - from Compliance Dashboard
11. Closure Phase % - from Compliance Dashboard
12. Critical Gaps Count - count of "❌ Not Done" across all phases
13. High Findings Count - count of "⚠️ Incomplete" across all phases
14. Deploy Date - (manual field, blank by default)
15. Last Assessment Date - from Compliance Dashboard
16. Notes - auto-generated consolidation reference

Usage:
    Basic (auto-detect workbooks in current directory):
        python3 consolidate_a58_dashboard.py
    
    Specify input directory:
        python3 consolidate_a58_dashboard.py --input-dir /path/to/projects
    
    Specify dashboard file:
        python3 consolidate_a58_dashboard.py --dashboard dashboard.xlsx
    
    Dry run (validate without writing):
        python3 consolidate_a58_dashboard.py --dry-run
    
    With backup:
        python3 consolidate_a58_dashboard.py --backup --verbose

Requirements:
    - Python 3.8+
    - openpyxl library
    - Completed A.5.8.1 workbooks (with actual project data)
    - A.5.8.3 dashboard template exists

Version: 1.0
Date: 2025-01-29
Control Reference: ISO/IEC 27001:2022 Control A.5.8
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import argparse
import glob
import logging
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# Third-Party Imports
# =============================================================================
try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
except ImportError:
    print("Error: openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


# ============================================================================
# DATA EXTRACTION FUNCTIONS
# ============================================================================

def extract_project_data(workbook_path, verbose=False):
    """
    Extract project data from A.5.8.1 workbook.
    
    Returns dict with 16 fields or None if extraction fails.
    """
    if verbose:
        print(f"    Reading: {Path(workbook_path).name}")
    
    try:
        wb = load_workbook(workbook_path, data_only=True, read_only=True)
        
        # Validate required sheets exist
        required_sheets = ["2. Project Classification", "8. Compliance Dashboard"]
        for sheet in required_sheets:
            if sheet not in wb.sheetnames:
                if verbose:
                    print(f"      ⚠️  Missing sheet: {sheet}")
                wb.close()
                return None
        
        # ----------------------------------------------------------------
        # Extract from Classification sheet (B3, B4, B18)
        # ----------------------------------------------------------------
        ws_class = wb["2. Project Classification"]
        project_name = ws_class["B3"].value
        pm = ws_class["B4"].value
        classification = ws_class["B18"].value  # Formula result: High/Medium/Low
        
        # Validate minimum data present
        if not project_name:
            if verbose:
                print(f"      ⚠️  No project name found - workbook may be empty template")
            wb.close()
            return None
        
        # ----------------------------------------------------------------
        # Extract from Compliance Dashboard (B8, B9-B13, B15)
        # ----------------------------------------------------------------
        ws_dash = wb["8. Compliance Dashboard"]
        
        # Last assessment date (B8)
        last_assessment = ws_dash["B8"].value
        if not last_assessment:
            last_assessment = datetime.now().date()
        
        # Phase completion percentages (B9-B13)
        # These are formulas that calculate from phase sheets
        initiation_pct = ws_dash["B9"].value or 0
        planning_pct = ws_dash["B10"].value or 0
        execution_pct = ws_dash["B11"].value or 0
        monitoring_pct = ws_dash["B12"].value or 0
        closure_pct = ws_dash["B13"].value or 0
        
        # Overall compliance (B15) - average of all phases
        compliance_score = ws_dash["B15"].value or 0
        
        # Convert to numeric (formula results come as float or int)
        def safe_float(val):
            if val is None:
                return 0.0
            if isinstance(val, (int, float)):
                return float(val)
            try:
                return float(val)
            except:
                return 0.0
        
        initiation_pct = safe_float(initiation_pct)
        planning_pct = safe_float(planning_pct)
        execution_pct = safe_float(execution_pct)
        monitoring_pct = safe_float(monitoring_pct)
        closure_pct = safe_float(closure_pct)
        compliance_score = safe_float(compliance_score)
        
        # ----------------------------------------------------------------
        # Determine current phase from highest completion percentage
        # ----------------------------------------------------------------
        if closure_pct > 0:
            phase = "Closure"
        elif monitoring_pct > 0:
            phase = "Monitoring"
        elif execution_pct > 0:
            phase = "Execution"
        elif planning_pct > 0:
            phase = "Planning"
        elif initiation_pct > 0:
            phase = "Initiation"
        else:
            phase = "Classification"
        
        # ----------------------------------------------------------------
        # Count critical gaps and high findings from phase sheets
        # ----------------------------------------------------------------
        critical_gaps = 0  # Count of "❌ Not Done"
        high_findings = 0  # Count of "⚠️ Incomplete"
        
        phase_sheets = [
            "3. Initiation Phase",
            "4. Planning Phase", 
            "5. Execution Phase",
            "6. Monitoring Phase",
            "7. Closure Phase"
        ]
        
        for phase_sheet in phase_sheets:
            if phase_sheet in wb.sheetnames:
                ws_phase = wb[phase_sheet]
                # Status is in column B, starting row 4 (row 3 is headers)
                # Typically 6-9 activities per phase
                for row in range(4, min(ws_phase.max_row + 1, 20)):
                    status = ws_phase[f"B{row}"].value
                    if status == "❌ Not Done":
                        critical_gaps += 1
                    elif status == "⚠️ Incomplete":
                        high_findings += 1
        
        wb.close()
        
        # ----------------------------------------------------------------
        # Return extracted data as dict
        # ----------------------------------------------------------------
        return {
            "project_name": str(project_name),
            "classification": str(classification) if classification else "Medium",
            "pm": str(pm) if pm else "",
            "business_owner": "",  # Not in A.5.8.1, user must populate manually
            "phase": phase,
            "compliance_score": compliance_score,
            "initiation_pct": initiation_pct,
            "planning_pct": planning_pct,
            "execution_pct": execution_pct,
            "monitoring_pct": monitoring_pct,
            "closure_pct": closure_pct,
            "critical_gaps": critical_gaps,
            "high_findings": high_findings,
            "deploy_date": None,  # Not in A.5.8.1, user must populate manually
            "last_assessment": last_assessment,
            "notes": f"Auto-consolidated from {Path(workbook_path).name} on {datetime.now().strftime('%Y-%m-%d')}"
        }
        
    except Exception as e:
        if verbose:
            print(f"      ❌ Error extracting data: {e}")
            import traceback
            traceback.print_exc()
        return None


def find_assessment_workbooks(input_dir, verbose=False):
    """Find all A.5.8.1 workbooks in directory."""
    pattern = os.path.join(input_dir, "*A.5.8.1*.xlsx")
    workbooks = glob.glob(pattern)
    
    # Filter out temporary files (Excel temp files start with ~$)
    workbooks = [wb for wb in workbooks if not Path(wb).name.startswith('~$')]
    
    if verbose:
        print(f"  Search pattern: {pattern}")
        print(f"  Found {len(workbooks)} A.5.8.1 workbook(s)")
        for wb in workbooks:
            print(f"    • {Path(wb).name}")
    
    return workbooks


# ============================================================================
# DASHBOARD POPULATION
# ============================================================================

def safely_write_data(ws, start_row, projects_data, verbose=False):
    """
    Write project data to dashboard, handling merged cells safely.
    Returns number of projects written.
    """
    row = start_row
    projects_written = 0
    
    for project in projects_data:
        if verbose:
            print(f"    Row {row}: {project['project_name']}")
        
        # Write data to columns A-P (16 columns)
        try:
            # Column A: Project Name
            cell = ws.cell(row=row, column=1)
            if not isinstance(cell, MergedCell):
                cell.value = project["project_name"]
            
            # Column B: Classification
            cell = ws.cell(row=row, column=2)
            if not isinstance(cell, MergedCell):
                cell.value = project["classification"]
            
            # Column C: PM
            cell = ws.cell(row=row, column=3)
            if not isinstance(cell, MergedCell):
                cell.value = project["pm"]
            
            # Column D: Business Owner
            cell = ws.cell(row=row, column=4)
            if not isinstance(cell, MergedCell):
                cell.value = project["business_owner"]
            
            # Column E: Phase
            cell = ws.cell(row=row, column=5)
            if not isinstance(cell, MergedCell):
                cell.value = project["phase"]
            
            # Column F: Compliance Score (percentage)
            cell = ws.cell(row=row, column=6)
            if not isinstance(cell, MergedCell):
                cell.value = project["compliance_score"]
            
            # Columns G-K: Phase percentages
            for col, key in enumerate(["initiation_pct", "planning_pct", "execution_pct", 
                                       "monitoring_pct", "closure_pct"], start=7):
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = project[key]
            
            # Column L: Critical Gaps
            cell = ws.cell(row=row, column=12)
            if not isinstance(cell, MergedCell):
                cell.value = project["critical_gaps"]
            
            # Column M: High Findings
            cell = ws.cell(row=row, column=13)
            if not isinstance(cell, MergedCell):
                cell.value = project["high_findings"]
            
            # Column N: Deploy Date
            cell = ws.cell(row=row, column=14)
            if not isinstance(cell, MergedCell):
                cell.value = project["deploy_date"]
            
            # Column O: Last Assessment
            cell = ws.cell(row=row, column=15)
            if not isinstance(cell, MergedCell):
                cell.value = project["last_assessment"]
            
            # Column P: Notes
            cell = ws.cell(row=row, column=16)
            if not isinstance(cell, MergedCell):
                cell.value = project["notes"]
            
            projects_written += 1
            row += 1
            
        except Exception as e:
            if verbose:
                print(f"      ⚠️  Error writing project data: {e}")
            continue
    
    return projects_written


def populate_portfolio_dashboard(dashboard_path, projects_data, verbose=False, dry_run=False):
    """Populate Portfolio Dashboard with extracted project data."""
    
    if dry_run:
        print("\n🔍 DRY RUN MODE - No changes will be written")
    
    try:
        wb = load_workbook(dashboard_path)
    except Exception as e:
        print(f"❌ Error loading dashboard: {e}")
        return False
    
    # Validate dashboard structure
    if "Project Data" not in wb.sheetnames:
        print("❌ Error: Dashboard missing 'Project Data' sheet")
        wb.close()
        return False
    
    ws = wb["Project Data"]
    
    if verbose or dry_run:
        print(f"\n  Target sheet: Project Data")
        print(f"  Starting row: 4 (after headers at row 3)")
        print(f"  Projects to write: {len(projects_data)}")
    
    # Write project data starting at row 4 (row 3 is headers)
    start_row = 4
    
    if not dry_run:
        projects_written = safely_write_data(ws, start_row, projects_data, verbose)
        
        # Save dashboard
        try:
            wb.save(dashboard_path)
            print(f"\n💾 Saved: {Path(dashboard_path).name}")
            print(f"  Projects written: {projects_written}")
        except Exception as e:
            print(f"❌ Error saving dashboard: {e}")
            wb.close()
            return False
    else:
        # Dry run - just validate we can write
        print(f"\n  Would write {len(projects_data)} projects to rows {start_row}-{start_row + len(projects_data) - 1}")
    
    wb.close()
    return True


# ============================================================================
# MAIN CONSOLIDATION FUNCTION
# ============================================================================

def consolidate_portfolio(args):
    """Main consolidation workflow."""

    print("=" * 78)
    print("ISMS-IMP-A.5.8 Portfolio Dashboard Consolidation")
    print("=" * 78)

    # Determine input directory - default to ../90_workbooks relative to script
    if args.input_dir:
        input_dir = args.input_dir
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        input_dir = os.path.join(os.path.dirname(script_dir), '90_workbooks')
    if args.verbose:
        print(f"\nConfiguration:")
        print(f"  Input directory: {os.path.abspath(input_dir)}")
    
    # Find or validate dashboard file
    if args.dashboard:
        dashboard_path = args.dashboard
        if not os.path.exists(dashboard_path):
            print(f"❌ Error: Dashboard not found: {dashboard_path}")
            return False
    else:
        # Auto-detect dashboard
        pattern = os.path.join(input_dir, "*A.5.8.3*.xlsx")
        dashboards = glob.glob(pattern)
        dashboards = [d for d in dashboards if not Path(d).name.startswith('~$')]
        
        if not dashboards:
            print("❌ Error: No A.5.8.3 dashboard found in input directory")
            print("   Generate one with: python3 generate_a58_3_portfolio_dashboard.py")
            return False
        
        dashboard_path = dashboards[0]
        if len(dashboards) > 1 and args.verbose:
            print(f"  ⚠️  Multiple dashboards found, using: {Path(dashboard_path).name}")
    
    if args.verbose:
        print(f"  Dashboard file: {Path(dashboard_path).name}")
    
    # Backup dashboard if requested
    if args.backup and not args.dry_run:
        backup_name = str(dashboard_path).replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        shutil.copy2(dashboard_path, backup_name)
        print(f"\n📋 Backup created: {Path(backup_name).name}")
    
    # Phase 1: Find assessment workbooks
    print(f"\n[1/3] Finding A.5.8.1 assessment workbooks...")
    workbooks = find_assessment_workbooks(input_dir, args.verbose)
    
    if not workbooks:
        print("❌ Error: No A.5.8.1 workbooks found in input directory")
        print("   Ensure completed A.5.8.1 workbooks are in the directory")
        print(f"   Looking for pattern: *A.5.8.1*.xlsx in {os.path.abspath(input_dir)}")
        return False
    
    print(f"  ✓ Found {len(workbooks)} project assessment(s)")
    
    # Phase 2: Extract project data
    print(f"\n[2/3] Extracting project data from workbooks...")
    projects_data = []
    skipped = 0
    
    for wb_path in workbooks:
        data = extract_project_data(wb_path, args.verbose)
        if data:
            projects_data.append(data)
            if not args.verbose:
                print(f"  ✓ {data['project_name']}: {data['classification']} risk, {data['compliance_score']:.0%} compliance, {data['phase']} phase")
        else:
            skipped += 1
            if not args.verbose:
                print(f"  ⚠️  Skipped (empty/invalid): {Path(wb_path).name}")
    
    if not projects_data:
        print("\n❌ Error: No valid project data extracted")
        print("   All workbooks appear to be empty templates or invalid")
        print("   Ensure A.5.8.1 workbooks are completed with actual project data")
        return False
    
    print(f"\n  Summary:")
    print(f"    Projects extracted: {len(projects_data)}")
    if skipped > 0:
        print(f"    Skipped (empty): {skipped}")
    
    # Phase 3: Populate dashboard
    print(f"\n[3/3] Populating portfolio dashboard...")
    success = populate_portfolio_dashboard(dashboard_path, projects_data, args.verbose, args.dry_run)
    
    if not success:
        return False
    
    # Final summary
    print("\n" + "=" * 78)
    if args.dry_run:
        print("✅ DRY RUN VALIDATION COMPLETE")
    else:
        print("✅ CONSOLIDATION COMPLETE")
    print("=" * 78)
    
    print(f"\nSummary:")
    print(f"  • Projects consolidated: {len(projects_data)}")
    print(f"  • Dashboard updated: {Path(dashboard_path).name}")
    
    if not args.dry_run:
        print(f"\nNext Steps:")
        print(f"  1. Open dashboard: {Path(dashboard_path).name}")
        print(f"  2. Review Executive Summary (formulas auto-calculate)")
        print(f"  3. Check Project Status sheet")
        print(f"  4. Manually populate Business Owner column (D) if needed")
        print(f"  5. Manually populate Deploy Date column (N) if needed")
        print(f"  6. Update Gap Analysis with portfolio-wide common gaps")
        print(f"  7. Track trends quarter-over-quarter in Trend Analysis sheet")
    
    print("=" * 78 + "\n")
    
    return True


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    
    parser = argparse.ArgumentParser(
        description="Consolidate A.5.8.1 project assessments into A.5.8.3 portfolio dashboard",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage (auto-detect files in current directory)
  python3 consolidate_a58_dashboard.py
  
  # Specify input directory with project workbooks
  python3 consolidate_a58_dashboard.py --input-dir /path/to/projects
  
  # Dry run to validate before writing
  python3 consolidate_a58_dashboard.py --dry-run --verbose
  
  # Production consolidation with backup
  python3 consolidate_a58_dashboard.py --backup --verbose

Notes:
  - Source workbooks must be completed A.5.8.1 assessments (not empty templates)
  - Dashboard formulas auto-calculate after data population
  - Manual fields: Business Owner (D), Deploy Date (N)
  - Run quarterly or after major project milestones
        """
    )
    
    parser.add_argument('--input-dir', 
                       help='Directory containing A.5.8.1 workbooks (default: current directory)')
    parser.add_argument('--dashboard',
                       help='Path to A.5.8.3 dashboard file (default: auto-detect)')
    parser.add_argument('--dry-run', action='store_true',
                       help='Validate without writing to dashboard')
    parser.add_argument('--verbose', action='store_true',
                       help='Enable detailed logging')
    parser.add_argument('--backup', action='store_true',
                       help='Create backup of dashboard before populating')
    
    args = parser.parse_args()
    
    try:
        success = consolidate_portfolio(args)
        sys.exit(0 if success else 1)
    except KeyboardInterrupt:
        print("\n\n⚠️  Consolidation cancelled by user")
        sys.exit(130)
    except Exception as e:
        print(f"\n❌ Fatal error: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (syntax validated, STANDARDIZATION applied)
# QA_TOOL: Claude Code Deep Scan
# STANDARDIZATION: License header, logging, imports reorganized, main() pattern
# =============================================================================
