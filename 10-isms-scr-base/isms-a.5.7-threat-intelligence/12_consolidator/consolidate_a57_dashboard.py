#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
ISMS-A.5.7 Threat Intelligence - Dashboard Consolidation Utility
Auto-detects workbooks in ../90_workbooks/ and consolidates data into dashboard.
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import os
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import load_workbook

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# Source workbook patterns and their assessment areas
SOURCE_WORKBOOKS = [
    ('A.5.7.1', 'A_5_7_1', 'TI Sources'),
    ('A.5.7.2', 'A_5_7_2', 'Collection & Analysis'),
    ('A.5.7.3', 'A_5_7_3', 'Integration'),
]

DASHBOARD_PATTERNS = ['A.5.7.4', 'A_5_7_4', 'Dashboard']


def find_workbook(directory, patterns):
    """Find workbook matching any of the patterns"""
    if isinstance(patterns, str):
        patterns = [patterns]
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            for pattern in patterns:
                if pattern in filename:
                    return os.path.join(directory, filename)
    return None


def find_workbooks_by_patterns(directory, pattern_list):
    """Find all workbooks matching pattern tuples"""
    found = []
    for patterns in pattern_list:
        if isinstance(patterns, tuple):
            patterns, alt_pattern, area = patterns[0], patterns[1], patterns[2]
            wb = find_workbook(directory, [patterns, alt_pattern])
            if wb:
                found.append((wb, area))
        else:
            wb = find_workbook(directory, patterns)
            if wb:
                found.append((wb, patterns))
    return found


def safely_write_data(ws, start_row, data):
    """Write data rows to worksheet"""
    count = 0
    for i, row_data in enumerate(data):
        for j, value in enumerate(row_data):
            if value is not None:
                ws.cell(row=start_row + i, column=j + 1, value=value)
        count += 1
    return count


def extract_gaps_from_workbook(filepath, assessment_area, gap_counter_start=1):
    """Extract gaps from source workbook"""
    gaps = []
    gap_counter = gap_counter_start
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if any(x in sheet_name.lower() for x in ['gap', 'finding', 'issue']):
                ws = wb[sheet_name]
                for row in range(8, min(ws.max_row + 1, 100)):
                    row_values = [ws.cell(row=row, column=col).value for col in range(1, 12)]
                    if not any(row_values):
                        continue
                    status = None
                    gap_desc = None
                    for val in row_values:
                        if val:
                            val_str = str(val)
                            if any(x in val_str for x in ['⚠️', '❌', 'Non-Compliant', 'Partial', 'Gap']):
                                status = val_str
                            elif len(val_str) > 30:
                                gap_desc = val_str
                    if status or gap_desc:
                        severity = 'High' if status and ('❌' in str(status) or 'Non-Compliant' in str(status)) else 'Medium'
                        gaps.append([
                            f'GAP-{assessment_area[:3].upper()}-{gap_counter:03d}',
                            assessment_area, os.path.basename(filepath),
                            row_values[0] if row_values[0] else 'System',
                            gap_desc[:200] if gap_desc else f'{assessment_area} gap',
                            status[:30] if status else 'Gap', 'Compliant', severity, 'Open',
                            f'Remediation required', None, None
                        ])
                        gap_counter += 1
        wb.close()
    except Exception as e:
        print(f"    ⚠️  Error: {e}")
    return gaps, gap_counter


def extract_evidence_from_workbook(filepath, assessment_area):
    """Extract evidence from source workbook"""
    evidence = []
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if 'evidence' in sheet_name.lower():
                ws = wb[sheet_name]
                for row in range(10, min(ws.max_row + 1, 50)):
                    evd_id = ws.cell(row=row, column=1).value
                    if evd_id and str(evd_id).startswith('EVD'):
                        evidence.append([evd_id, assessment_area, os.path.basename(filepath), None,
                            ws.cell(row=row, column=2).value, ws.cell(row=row, column=3).value,
                            ws.cell(row=row, column=4).value, ws.cell(row=row, column=5).value,
                            ws.cell(row=row, column=6).value, ws.cell(row=row, column=7).value,
                            ws.cell(row=row, column=8).value, ws.cell(row=row, column=9).value])
        wb.close()
    except Exception as e:
        print(f"    ⚠️  Error: {e}")
    return evidence


def generate_risks_from_gaps(gaps):
    """Generate risks from high severity gaps"""
    risks = []
    for i, gap in enumerate([g for g in gaps if g[7] == 'High'], 1):
        risks.append([f'RISK-{i:03d}', gap[0], 'Security', f'Risk: {gap[4][:80]}',
            gap[3], 'High', 'High', 'High', 'Open', f'Mitigate {gap[0]}', None, None, None])
    return risks


def generate_remediation_from_gaps(gaps):
    """Generate remediation actions from gaps"""
    remediation = []
    today = datetime.now()
    for i, gap in enumerate(gaps, 1):
        days = 30 if gap[7] == 'High' else 60
        remediation.append([f'REM-{i:03d}', gap[0], gap[1], gap[3], gap[9] or 'Remediation required',
            'Critical' if gap[7] == 'High' else 'High', 'Planned', None,
            today.strftime('%Y-%m-%d'), (today + timedelta(days=days)).strftime('%Y-%m-%d'),
            None, '0%', None])
    return remediation


def consolidate_dashboard():
    """Main consolidation function"""
    print("=" * 80)
    print("ISMS-A.5.7 Threat Intelligence - Dashboard Consolidation")
    print("=" * 80)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    workbooks_dir = os.path.join(os.path.dirname(script_dir), '90_workbooks')

    if not os.path.exists(workbooks_dir):
        print(f"❌ Workbooks directory not found: {workbooks_dir}")
        return False

    print(f"\nWorkbooks directory: {workbooks_dir}")

    # Find dashboard
    dashboard_path = find_workbook(workbooks_dir, DASHBOARD_PATTERNS)
    if not dashboard_path:
        print("❌ Dashboard not found")
        return False
    print(f"Dashboard: {os.path.basename(dashboard_path)}")

    # Find source workbooks
    print("\n[1/5] Finding source workbooks...")
    sources = find_workbooks_by_patterns(workbooks_dir, SOURCE_WORKBOOKS)
    for path, area in sources:
        print(f"  ✓ {area}: {os.path.basename(path)}")

    # Extract data
    print("\n[2/5] Extracting gaps...")
    all_gaps = []
    gap_counter = 1
    for filepath, area in sources:
        gaps, gap_counter = extract_gaps_from_workbook(filepath, area, gap_counter)
        print(f"  • {area}: {len(gaps)} gaps")
        all_gaps.extend(gaps)

    print("\n[3/5] Extracting evidence...")
    all_evidence = []
    for filepath, area in sources:
        evidence = extract_evidence_from_workbook(filepath, area)
        print(f"  • {area}: {len(evidence)} items")
        all_evidence.extend(evidence)

    print("\n[4/5] Generating risks and remediation...")
    all_risks = generate_risks_from_gaps(all_gaps)
    all_remediation = generate_remediation_from_gaps(all_gaps)
    print(f"  • Risks: {len(all_risks)}")
    print(f"  • Remediation: {len(all_remediation)}")

    print("\n[5/5] Writing to dashboard...")
    try:
        wb = load_workbook(dashboard_path)
        for sheet_name, data, start_row in [
            ('Gap Analysis', all_gaps, 14), ('Gap_Analysis', all_gaps, 14),
            ('Risk Register', all_risks, 21), ('Risk_Register', all_risks, 21),
            ('Remediation Roadmap', all_remediation, 37), ('Remediation_Roadmap', all_remediation, 37),
            ('Evidence Register', all_evidence, 14), ('Evidence_Register', all_evidence, 14),
        ]:
            if sheet_name in wb.sheetnames and data:
                count = safely_write_data(wb[sheet_name], start_row, data)
                print(f"  ✓ {sheet_name}: {count} entries")
        wb.save(dashboard_path)
        print(f"\n💾 Saved: {os.path.basename(dashboard_path)}")
        wb.close()
    except Exception as e:
        print(f"  ❌ Error: {e}")
        return False

    print("\n" + "=" * 80)
    print("✅ CONSOLIDATION COMPLETE")
    print(f"  Gaps: {len(all_gaps)} | Risks: {len(all_risks)} | Remediation: {len(all_remediation)} | Evidence: {len(all_evidence)}")
    print("=" * 80 + "\n")
    return True


if __name__ == "__main__":
    success = consolidate_dashboard()
    sys.exit(0 if success else 1)

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (syntax validated, STANDARDIZATION applied)
# QA_TOOL: Claude Code Deep Scan
# STANDARDIZATION: License header, logging, imports reorganized, main() pattern
# =============================================================================
