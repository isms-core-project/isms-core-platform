#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.7.4 - Threat Intelligence Effectiveness Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.7: Threat Intelligence
Assessment Domain 4 of 5: Compliance Dashboard (Consolidated Metrics)

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific assessment workbook structures, compliance
thresholds, and reporting requirements.

Key customization areas:
1. Input workbook schemas (match actual sheet structures from A.5.7.1-3)
2. Compliance scoring formulas (adapt to your risk tolerance)
3. Dashboard layout and KPIs (align with executive reporting needs)
4. Trend analysis periods (match your reporting cycles)
5. Threshold values for compliance status (based on audit requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.7 Threat Intelligence Framework

⚠️ CRITICAL: This script consolidates data from multiple input workbooks.
You MUST analyze the actual structure of each input workbook before use:
- A.5.7.1 Sources Assessment (15 sheets)
- A.5.7.2 Collection & Analysis Assessment (12 sheets)
- A.5.7.3 Integration & Distribution Assessment (13 sheets)

DO NOT assume sheet names or column structures - validate against actual files.

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel dashboard that consolidates
assessment data from Control A.5.7 Domains 1-3 into executive-level metrics,
compliance status, and trend analysis for management oversight and audit
readiness.

**Purpose:**
Enables executive oversight of threat intelligence program effectiveness by
consolidating operational assessments (sources, collection/analysis,
integration/distribution) into strategic metrics, compliance status, and
actionable insights for continuous improvement.

**Assessment Scope:**
- Overall Control A.5.7 compliance status and scoring
- Source portfolio health and coverage effectiveness (from A.5.7.1)
- Collection/analysis workflow performance (from A.5.7.2)
- Integration/distribution effectiveness (from A.5.7.3)
- VulnerabilityThreatLink (VTL) workflow health (cross-control integration)
- CVSS accuracy and quality metrics (critical for VTL)
- MITRE ATT&CK coverage across the threat intelligence lifecycle
- Key performance indicators (KPIs) and trend analysis
- Gap summary and prioritized remediation roadmap
- Audit readiness status and evidence completeness
- Executive summary for management briefings

**Generated Dashboard Structure (10 Sheets):**
1. Executive_Summary - High-level compliance status and KPIs
2. Source_Portfolio_Health - Consolidated metrics from A.5.7.1
3. Collection_Analysis_Performance - Consolidated metrics from A.5.7.2
4. Integration_Distribution_Effectiveness - Consolidated metrics from A.5.7.3
5. VTL_Workflow_Health - VulnerabilityThreatLink integration status (CRITICAL)
6. CVSS_Quality_Metrics - CVSS accuracy and completeness tracking
7. MITRE_Coverage_Heatmap - ATT&CK coverage across lifecycle stages
8. KPI_Dashboard - Key performance indicators with trend visualization
9. Gap_Remediation_Roadmap - Prioritized improvement plan
10. Audit_Readiness - Evidence completeness and compliance validation

**Key Features:**
- Automated data consolidation from three assessment workbooks
- Compliance scoring with weighted metrics across domains
- VTL workflow health tracking (A.5.7 ↔ A.8.8 integration)
- CVSS quality dashboard for audit evidence
- MITRE ATT&CK heatmap aggregation across source/collection/distribution
- Trend analysis with historical comparison (if prior assessments available)
- Executive summary with traffic light status indicators
- Gap prioritization based on compliance impact and remediation effort
- Conditional formatting for visual status communication
- Protected formulas with external workbook references
- Evidence completeness checklist for audit readiness

**VulnerabilityThreatLink (VTL) Dashboard (CRITICAL):**
Sheet 5 (VTL_Workflow_Health) provides executive visibility into the critical
integration between Control A.5.7 (Threat Intelligence) and Control A.8.8
(Vulnerability Management):

VTL Metrics Displayed:
- CVSS coverage rate (% of threat intel with CVSS scores)
- CVSS accuracy score (validation against authoritative sources)
- Emergency patch trigger rate (exploited CVEs auto-escalated)
- VTL workflow latency (threat detection → patch deployment time)
- Remediation feedback loop health (vuln team → threat intel)

This dashboard demonstrates operational security value to auditors and executives.

**Integration:**
This dashboard consolidates data from:
- A.5.7.1 Threat Intelligence Sources Assessment (15 sheets)
- A.5.7.2 Intelligence Collection & Analysis Assessment (12 sheets)
- A.5.7.3 Intelligence Integration & Distribution Assessment (13 sheets)

Data is referenced via external workbook formulas for real-time updates.

Feeds into:
- A.5.7.5 Standalone Compliance Dashboard (executive reporting without dependencies)
- Management review meetings and board reporting
- Audit evidence packages for ISO 27001 certification
- Continuous improvement planning and budget justification

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

Input Files Required (must be in same directory):
    - ISMS_A_5_7_1_Sources_Assessment_YYYYMMDD.xlsx
    - ISMS_A_5_7_2_Collection_Analysis_Assessment_YYYYMMDD.xlsx
    - ISMS_A_5_7_3_Integration_Distribution_Assessment_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a57_4_compliance_dashboard.py

Advanced Usage:
    # Generate with custom input directory
    python3 generate_a57_4_compliance_dashboard.py --input /path/to/assessments
    
    # Generate with specific date suffix
    python3 generate_a57_4_compliance_dashboard.py --date 20250115
    
    # Specify input file dates (if different from output date)
    python3 generate_a57_4_compliance_dashboard.py --sources-date 20250110 \
        --collection-date 20250112 --integration-date 20250114

Output:
    File: ISMS_A_5_7_4_Effectiveness_Dashboard_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)
    Sheets: 10 worksheets (see structure above)

Prerequisites:
    1. Complete A.5.7.1 Sources Assessment workbook
    2. Complete A.5.7.2 Collection & Analysis Assessment workbook
    3. Complete A.5.7.3 Integration & Distribution Assessment workbook
    4. All input workbooks in same directory as this script

Post-Generation Steps:
    1. Verify all external workbook references are functioning
    2. Review consolidated compliance scores and status
    3. Analyze VTL workflow health metrics (Sheet 5 - CRITICAL)
    4. Review CVSS quality metrics for accuracy and completeness
    5. Examine MITRE ATT&CK coverage heatmap for gaps
    6. Prioritize gap remediation based on compliance impact
    7. Prepare executive summary narrative (customize Sheet 1 text)
    8. Validate audit evidence completeness (Sheet 10)
    9. Present to management for review and approval
    10. Archive dashboard with assessment workbooks for audit trail
    11. Update A.5.7.5 Standalone Dashboard if needed

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.7
Assessment Domain:    4 of 5 (Compliance Dashboard - Consolidated Metrics)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.7: Threat Intelligence Policy (Governance)
    - ISMS-IMP-A.5.7.4: Threat Intelligence Effectiveness Dashboard Implementation Guide
    - ISMS-IMP-A.5.7.1: Threat Intelligence Sources Assessment (Domain 1)
    - ISMS-IMP-A.5.7.2: Intelligence Collection & Analysis Assessment (Domain 2)
    - ISMS-IMP-A.5.7.3: Intelligence Integration & Distribution Assessment (Domain 3)
    - ISMS-IMP-A.5.7.5: Standalone Compliance Dashboard (Executive Reporting)
    - ISMS-POL-A.8.8: Vulnerability Management Policy (VTL integration)
    - ISO 27002:2022 Implementation Guidance for Control A.5.7
    - MITRE ATT&CK Framework (Enterprise v15.1)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
- Added dedicated VTL_Workflow_Health sheet (Sheet 5) for A.5.7↔A.8.8 integration
    - Enhanced CVSS quality metrics dashboard (Sheet 6)
    - Expanded MITRE ATT&CK coverage heatmap with lifecycle stages
    - Improved gap prioritization with compliance impact scoring
    - Added audit readiness checklist with evidence completeness tracking
    - Updated executive summary with VTL workflow status
    - Enhanced conditional formatting for visual status communication
    - Added trend analysis support (if historical assessments available)

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Dashboard Purpose:**
This dashboard is designed for EXECUTIVE OVERSIGHT and AUDIT READINESS, not
operational use. It provides:
- Strategic view of threat intelligence program health
- Compliance status against ISO 27001:2022 requirements
- Evidence of operational security value (VTL workflow)
- Prioritized improvement roadmap for management decisions

For operational metrics, users should reference the source assessments directly.

**External Workbook References:**
This dashboard uses Excel external workbook references to consolidate data.
This design ensures:
- Real-time updates when source assessments are modified
- No data duplication or synchronization issues
- Audit trail preservation (source data remains in original workbooks)

CRITICAL: External references break if source workbooks are moved, renamed,
or deleted. Always keep all A.5.7 assessment workbooks in the same directory.

**VTL Workflow Dashboard (AUDIT CRITICAL):**
Sheet 5 (VTL_Workflow_Health) is the most scrutinized dashboard by auditors.
It demonstrates that threat intelligence provides OPERATIONAL VALUE beyond
compliance checkbox exercises.

Auditors will specifically verify:
- CVSS coverage and accuracy metrics (enables automation)
- Emergency patching trigger rates (threat intel drives action)
- Workflow latency metrics (time from threat detection to remediation)
- Bi-directional feedback (vulnerability team communicates back to threat intel)

Incomplete or poor VTL metrics suggest "security theater" rather than effective
threat intelligence operations.

**CVSS Quality Metrics (VTL Dependency):**
Sheet 6 (CVSS_Quality_Metrics) consolidates CVSS accuracy from:
- Source portfolio CVSS support (A.5.7.1)
- Collection/analysis CVSS enrichment (A.5.7.2)
- Integration CVSS validation (A.5.7.3)

Target: >90% CVSS accuracy across all sources and enrichment processes.
Below 85% accuracy degrades vulnerability management automation quality.

**Compliance Scoring Methodology:**
Overall Control A.5.7 compliance is calculated as weighted average:
- Source Portfolio (A.5.7.1): 30% weight
- Collection & Analysis (A.5.7.2): 35% weight
- Integration & Distribution (A.5.7.3): 35% weight

Within each domain, specific metrics are weighted by compliance criticality.
Customize weights in script to match your risk profile and audit focus areas.

**MITRE ATT&CK Coverage Heatmap:**
Sheet 7 aggregates ATT&CK coverage across lifecycle stages:
- Source coverage: Techniques covered by threat intelligence sources
- Analysis coverage: Techniques analysts can detect/investigate
- Integration coverage: Techniques with deployed detection/response

Target: >80% coverage of techniques relevant to your threat model.
Gaps indicate blind spots in threat detection capabilities.

**Audit Considerations:**
This dashboard serves as primary audit evidence for Control A.5.7 compliance.
Auditors will verify:
- Comprehensive coverage across all domains (sources, collection, integration)
- Evidence of operational effectiveness (VTL workflow health)
- Regular review and continuous improvement (trend analysis)
- Management oversight and approval (sign-offs on Sheet 1)
- Gap remediation planning and execution (Sheet 9)

Incomplete dashboards may result in audit findings or corrective action requests.

**Data Protection:**
Dashboard consolidates sensitive operational details:
- Threat intelligence program gaps and deficiencies
- VTL workflow metrics revealing vulnerability response capabilities
- CVSS accuracy metrics indicating detection/response quality
- Gap remediation plans exposing security improvement priorities

Classification: CONFIDENTIAL - Management Use Only. Limit distribution to
executives, ISMS team, and auditors.

**Maintenance:**
Update dashboard:
- Quarterly: After completing A.5.7.1-3 assessments each quarter
- Semi-annually: Update compliance thresholds based on audit feedback
- Annually: Complete reassessment with trend analysis vs. prior year
- Ad-hoc: When significant program changes or audit preparation needed

**Quality Assurance:**
Have CISO, threat intelligence lead, and ISMS manager review dashboard before
management presentation or audit submission. Cross-validate VTL metrics with
Control A.8.8 vulnerability management dashboard data.

**Historical Trend Analysis:**
If you maintain historical assessment data, this dashboard can display trends:
- Compliance score improvement over time
- CVSS quality improvement trajectory
- MITRE ATT&CK coverage expansion
- Gap reduction progress

Archive prior quarter dashboards to enable trend analysis functionality.

**Regulatory Alignment:**
Some regulations require specific threat intelligence program metrics:
- Financial sector: Fraud intelligence effectiveness metrics
- Healthcare: Healthcare-specific threat intelligence coverage
- Critical infrastructure: CISA information sharing participation
- Data protection: Privacy-preserving threat intelligence sharing metrics

Customize dashboard KPIs to include regulatory-specific requirements.

**Executive Communication:**
Sheet 1 (Executive_Summary) is designed for board/executive presentation.
Customize narrative text to explain:
- Overall program health in business terms
- Key achievements and improvements
- Critical gaps requiring investment/attention
- Compliance status and audit readiness
- VTL workflow value (quantified security improvements)

Avoid excessive technical jargon - focus on risk reduction and business value.

**Broken External References:**
If external workbook references break (moved files, renamed sources):
1. Check source workbook filenames and locations
2. Update external reference paths in dashboard formulas
3. Regenerate dashboard with corrected input file paths
4. Validate all metrics display correctly after reference fixes

Consider using relative paths or documenting expected directory structure.

**Business Impact:**
This dashboard enables data-driven decisions about:
- Threat intelligence budget allocation and ROI justification
- Resource allocation (analyst hiring, tool procurement)
- Gap remediation prioritization (highest compliance/security impact)
- Vendor selection and contract renewals (source portfolio optimization)
- Integration investments (improving VTL automation)

Poor dashboard quality leads to uninformed decisions and wasted investments.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)




# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.7.4"
WORKBOOK_NAME = "Threat Intelligence Effectiveness Dashboard"
CONTROL_ID = "A.5.7"
CONTROL_NAME = "Threat Intelligence"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching specification."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches ISMS-IMP-A.5.7.4 specification (12 sheets)
    sheets = [
        "Executive_Summary",
        "Program_KPIs",
        "Source_Portfolio",
        "Intelligence_Operations",
        "Integration_Status",
        "Stakeholder_Engagement",
        "Trend_Analysis",
        "Risk_Indicators",
        "Compliance_Evidence",
        "Monthly_Report",
        "Quarterly_Report",
        "Metadata",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="8FAADC", end_color="8FAADC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "formula_cell": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "dashboard_metric": {
            "font": Font(name="Calibri", size=18, bold=True),
            "fill": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
    }
    return styles


def get_border():
    """Create new border object (avoid shared object warnings)."""
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATION FUNCTIONS
# ============================================================================

def setup_validations():
    """Create all data validation dropdowns."""
    validations = {}

    # KPI Category
    validations['kpi_category'] = DataValidation(
        type="list",
        formula1='"Source_Management,Collection,Analysis,Production,Integration,Dissemination,Framework_Currency,Quality_Assurance,Emergency_Response,Impact_Measurement,Risk_Management,Incident_Response"',
        allow_blank=False
    )
    
    # Measurement Frequency
    validations['measurement_frequency'] = DataValidation(
        type="list",
        formula1='"Daily,Weekly,Monthly,Quarterly"',
        allow_blank=False
    )
    
    # Status
    validations['status'] = DataValidation(
        type="list",
        formula1='"On_Track,Caution,At_Risk"',
        allow_blank=False
    )
    
    # Trend
    validations['trend'] = DataValidation(
        type="list",
        formula1='"Improving,Stable,Declining"',
        allow_blank=True
    )
    
    # Risk Status
    validations['risk_status'] = DataValidation(
        type="list",
        formula1='"Normal,Warning,Critical"',
        allow_blank=False
    )
    
    # Evidence Status
    validations['evidence_status'] = DataValidation(
        type="list",
        formula1='"Current,Outdated,Missing"',
        allow_blank=False
    )
    
    return validations


# ============================================================================
# SECTION 3: EXTERNAL REFERENCE DOCUMENTATION
# ============================================================================

def get_external_reference_note():
    """
    Return instruction text for updating external references.
    CRITICAL: Users must update these paths after generation.
    """
    note = """
⚠️ EXTERNAL REFERENCE SETUP REQUIRED ⚠️

This dashboard uses external references to aggregate data from:
- ISMS-IMP-A.5.7.1 (Threat Intelligence Sources)
- ISMS-IMP-A.5.7.2 (Collection & Analysis)
- ISMS-IMP-A.5.7.3 (Integration & Distribution)

AFTER GENERATION, you must:
1. Place all 5.7 workbooks in the same directory
2. Open this dashboard workbook
3. Go to Data → Edit Links (or Data → Connections)
4. Update source file paths to match your directory structure

Example paths to update:
  [ISMS-IMP-A.5.7.1.xlsx]Source_Inventory!A:B
  [ISMS-IMP-A.5.7.2.xlsx]Intelligence_Production!A:C
  [ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment!A:T

When opening this file, Excel will prompt: "Update Links?" → Click UPDATE.

For formulas with external references, see Program_KPIs sheet column "Data_Source".
"""
    return note



# ============================================================================
# SECTION 3A: CVSS CALCULATION FUNCTIONS (NEW IN V1.0)
# ============================================================================

def calculate_cvss_adoption_rate():
    """
    Calculate CVSS 4.0 adoption rate from VTL records.
    
    Returns Excel formula string that references IMP-A.5.7.2 Sheet 8.
    Formula counts records with CVSS_Version="4.0" divided by total VTL records.
    """
    formula = (
        '=IF(COUNTA([ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Threat_Link!A:A)-4=0,0,'
        'COUNTIF([ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Threat_Link!E:E,"4.0")/'
        '(COUNTA([ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Threat_Link!A:A)-4)*100)'
    )
    return formula


def calculate_avg_cvss_prevented():
    """
    Calculate average CVSS score from prevented incidents.
    
    Returns Excel formula string that references IMP-A.5.7.3 Sheet 7.
    """
    formula = (
        '=IFERROR(AVERAGE([ISMS-IMP-A.5.7.3.xlsx]Prevention_Tracking!F:F),0)'
    )
    return formula


def calculate_high_cvss_active_exploitation():
    """
    Count VTL records with CVSS>=7.0 AND active/mass exploitation AND open status.
    
    CRITICAL: This is an emergency indicator. Result must be 0.
    Returns Excel formula string.
    """
    formula = (
        '=SUMPRODUCT('
        '([ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Threat_Link!F:F>=7)*'
        '(([ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Threat_Link!L:L="Active Exploitation")+'
        '([ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Threat_Link!L:L="Mass Exploitation"))*'
        '([ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Threat_Link!M:M="Open"))'
    )
    return formula


def calculate_cvss_accuracy_overall():
    """
    Calculate overall CVSS accuracy rate across all sources.
    
    Returns Excel formula string that references IMP-A.5.7.1 Sheet 14.
    """
    formula = (
        '=IFERROR(AVERAGE([ISMS-IMP-A.5.7.1.xlsx]Source_Performance_Validation!O:O),0)'
    )
    return formula


def calculate_cvss_quantification_rate():
    """
    Calculate % of vulnerability-related risks with CVSS scores.
    
    Returns Excel formula string that references IMP-A.5.7.3 Sheet 13.
    Target: 100% for Clause 6.1 compliance.
    """
    formula = (
        '=IF(COUNTIF([ISMS-IMP-A.5.7.3.xlsx]Risk_Assessment_Updates!F:F,"Vulnerability")=0,100,'
        'COUNTIFS([ISMS-IMP-A.5.7.3.xlsx]Risk_Assessment_Updates!F:F,"Vulnerability",'
        '[ISMS-IMP-A.5.7.3.xlsx]Risk_Assessment_Updates!H:H,"<>")/'
        'COUNTIF([ISMS-IMP-A.5.7.3.xlsx]Risk_Assessment_Updates!F:F,"Vulnerability")*100)'
    )
    return formula


def calculate_incident_cvss_context_rate():
    """
    Calculate % of vulnerability incidents with CVSS context.
    
    Returns Excel formula string that references IMP-A.5.7.3 Sheet 14.
    """
    formula = (
        '=IF(COUNTIF([ISMS-IMP-A.5.7.3.xlsx]Incident_TI_Integration!D:D,"Exploitation")=0,100,'
        'COUNTIFS([ISMS-IMP-A.5.7.3.xlsx]Incident_TI_Integration!D:D,"Exploitation",'
        '[ISMS-IMP-A.5.7.3.xlsx]Incident_TI_Integration!H:H,"<>")/'
        'COUNTIF([ISMS-IMP-A.5.7.3.xlsx]Incident_TI_Integration!D:D,"Exploitation")*100)'
    )
    return formula


def get_cvss_severity_formulas():
    """
    Return formulas for counting prevented incidents by CVSS severity.
    
    Returns dict with Critical/High/Medium/Low category formulas.
    """
    formulas = {
        "Critical": '=COUNTIF([ISMS-IMP-A.5.7.3.xlsx]Prevention_Tracking!F:F,">=9")',
        "High": '=COUNTIFS([ISMS-IMP-A.5.7.3.xlsx]Prevention_Tracking!F:F,">=7",[ISMS-IMP-A.5.7.3.xlsx]Prevention_Tracking!F:F,"<9")',
        "Medium": '=COUNTIFS([ISMS-IMP-A.5.7.3.xlsx]Prevention_Tracking!F:F,">=4",[ISMS-IMP-A.5.7.3.xlsx]Prevention_Tracking!F:F,"<7")',
        "Low": '=COUNTIFS([ISMS-IMP-A.5.7.3.xlsx]Prevention_Tracking!F:F,">0",[ISMS-IMP-A.5.7.3.xlsx]Prevention_Tracking!F:F,"<4")',
    }
    return formulas


def get_cvss_kpis():
    """
    Return CVSS-related KPIs to add to Program_KPIs sheet.
    
    Returns list of tuples matching the KPI structure.
    """
    cvss_kpis = [
        # CVSS Framework Health
        ("KPI-CVSS-001", "Framework_Currency", "CVSS 4.0 Adoption Rate",
         "% of VTL records using CVSS 4.0", "Percentage", "Monthly",
         calculate_cvss_adoption_rate(), "75%", "60%", "50%"),
        
        ("KPI-CVSS-002", "Quality_Assurance", "CVSS Accuracy Rate (All Sources)",
         "Average CVSS accuracy across all TI sources", "Percentage", "Monthly",
         calculate_cvss_accuracy_overall(), "90%", "85%", "80%"),
        
        ("KPI-CVSS-003", "Emergency_Response", "High CVSS + Active Exploitation (Open)",
         "Count of CVSS>=7.0 with active exploitation still unpatched", "Count", "Daily",
         calculate_high_cvss_active_exploitation(), "0", "0", "1"),
        
        ("KPI-CVSS-004", "Impact_Measurement", "Avg CVSS (Prevented Incidents)",
         "Average CVSS score of prevented incidents", "Score (0-10)", "Monthly",
         calculate_avg_cvss_prevented(), "N/A", "N/A", "N/A"),
        
        ("KPI-RISK-001", "Risk_Management", "Risk Updates with CVSS Quantification",
         "% of vulnerability risks with CVSS scores (Clause 6.1)", "Percentage", "Monthly",
         calculate_cvss_quantification_rate(), "100%", "95%", "90%"),
        
        ("KPI-INC-001", "Incident_Response", "Incidents with CVSS Context",
         "% of vulnerability incidents with CVSS data", "Percentage", "Monthly",
         calculate_incident_cvss_context_rate(), "100%", "90%", "80%"),
    ]
    
    return cvss_kpis


def add_cvss_metrics_to_executive_summary(ws, styles, start_row=20):
    """
    Add CVSS metrics section to Executive Summary sheet.
    
    NEW IN V1.0: CVSS Framework Health tracking.
    """
    row = start_row
    
    # Section header
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "CVSS FRAMEWORK HEALTH"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 25
    
    row += 1
    
    # Column headers
    headers = ["Metric", "Target", "Actual", "Status"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    row += 1
    
    # CVSS 4.0 Adoption Rate
    ws[f"A{row}"] = "CVSS 4.0 Adoption Rate"
    ws[f"B{row}"] = "≥75% by Q4 2026"
    ws[f"C{row}"] = calculate_cvss_adoption_rate()
    ws[f"C{row}"].number_format = '0.0"%"'
    ws[f"D{row}"] = f'=IF(C{row}>=75,"✅ Met",IF(C{row}>=60,"🔄 In Progress","⚠️ Behind"))'
    
    for col in [1, 2, 3, 4]:
        ws.cell(row=row, column=col).border = get_border()
    
    row += 1
    
    # Avg CVSS (Prevented)
    ws[f"A{row}"] = "Avg CVSS Score (Prevented CVEs)"
    ws[f"B{row}"] = "N/A (tracking only)"
    ws[f"C{row}"] = calculate_avg_cvss_prevented()
    ws[f"C{row}"].number_format = '0.0'
    ws[f"D{row}"] = "-"
    
    for col in [1, 2, 3, 4]:
        ws.cell(row=row, column=col).border = get_border()
    
    row += 1
    
    # High CVSS + Active Exploitation (CRITICAL INDICATOR)
    ws[f"A{row}"] = "High CVSS + Active Exploitation (Open)"
    ws[f"B{row}"] = "0 (all patched)"
    ws[f"C{row}"] = calculate_high_cvss_active_exploitation()
    ws[f"C{row}"].number_format = '0'
    ws[f"D{row}"] = f'=IF(C{row}=0,"✅ Compliant","❌ EMERGENCY ACTION REQUIRED")'
    
    # CRITICAL: Red background if count > 0
    ws.conditional_formatting.add(
        f'C{row}:D{row}',
        CellIsRule(operator='greaterThan', formula=['0'], 
                  fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                  font=Font(color="FFFFFF", bold=True))
    )
    
    for col in [1, 2, 3, 4]:
        ws.cell(row=row, column=col).border = get_border()
    
    row += 1
    
    # CVSS Accuracy Rate
    ws[f"A{row}"] = "CVSS Accuracy Rate (All Sources)"
    ws[f"B{row}"] = "≥90%"
    ws[f"C{row}"] = calculate_cvss_accuracy_overall()
    ws[f"C{row}"].number_format = '0.0"%"'
    ws[f"D{row}"] = f'=IF(C{row}>=90,"✅ Met",IF(C{row}>=85,"⚠️ Caution","❌ Missed"))'
    
    for col in [1, 2, 3, 4]:
        ws.cell(row=row, column=col).border = get_border()
    
    row += 2  # Skip a row
    
    # CVSS Severity Distribution (data for chart)
    ws[f"A{row}"] = "CVSS Severity Distribution (Prevented Incidents)"
    ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
    
    row += 1
    
    # Chart data
    severity_formulas = get_cvss_severity_formulas()
    chart_data = [
        ("Critical (9.0-10.0)", severity_formulas["Critical"]),
        ("High (7.0-8.9)", severity_formulas["High"]),
        ("Medium (4.0-6.9)", severity_formulas["Medium"]),
        ("Low (0.1-3.9)", severity_formulas["Low"]),
    ]
    
    ws[f"A{row}"] = "Severity Range"
    ws[f"B{row}"] = "Count"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].font = Font(bold=True)
    row += 1
    
    for severity_label, formula in chart_data:
        ws[f"A{row}"] = severity_label
        ws[f"B{row}"] = formula
        ws[f"B{row}"].number_format = '0'
        row += 1
    
    # Note for manual chart creation
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = f"📊 CREATE PIE CHART: Select A{row-5}:B{row-2}, Insert > Pie Chart, Title: 'Prevented Incidents by CVSS Severity'"
    ws[f"A{row}"].font = Font(name="Calibri", size=9, italic=True, color="0000FF")
    ws[f"A{row}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")


def add_cvss_quantification_tracking(ws, styles, start_row=30):
    """
    Add CVSS quantification tracking to Risk_Indicators sheet.
    
    NEW IN V1.0: Track Clause 6.1 compliance.
    """
    row = start_row
    
    # Section header
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "CVSS-BASED RISK QUANTIFICATION (ISO 27001 Clause 6.1)"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="8FAADC", end_color="8FAADC", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 25
    
    row += 1
    
    # Column headers
    headers = ["Quarter", "Total Risk Updates", "Vulnerability-Related", "Risks with CVSS", 
               "CVSS Quantification Rate", "Avg CVSS (Risks)", "Status"]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    row += 1
    
    # Current quarter (formula-based)
    ws[f"A{row}"] = f'=TEXT(TODAY(),"YYYY-\QQ")'
    ws[f"B{row}"] = '=COUNTA([ISMS-IMP-A.5.7.3.xlsx]Risk_Assessment_Updates!A:A)-3'
    ws[f"C{row}"] = '=COUNTIF([ISMS-IMP-A.5.7.3.xlsx]Risk_Assessment_Updates!F:F,"Vulnerability")'
    ws[f"D{row}"] = '=COUNTIFS([ISMS-IMP-A.5.7.3.xlsx]Risk_Assessment_Updates!F:F,"Vulnerability",[ISMS-IMP-A.5.7.3.xlsx]Risk_Assessment_Updates!H:H,"<>")'
    ws[f"E{row}"] = f'=IF(C{row}=0,"N/A",D{row}/C{row})'
    ws[f"E{row}"].number_format = '0.0"%"'
    ws[f"F{row}"] = '=AVERAGEIFS([ISMS-IMP-A.5.7.3.xlsx]Risk_Assessment_Updates!H:H,[ISMS-IMP-A.5.7.3.xlsx]Risk_Assessment_Updates!H:H,"<>")'
    ws[f"F{row}"].number_format = '0.0'
    ws[f"G{row}"] = f'=IF(E{row}=1,"✅ Compliant",IF(E{row}>=0.9,"⚠️ Needs Attention","❌ Non-Compliant"))'
    
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = get_border()
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'E{row}:E{row+5}',
        CellIsRule(operator='lessThan', formula=['0.9'], 
                  fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f'E{row}:E{row+5}',
        CellIsRule(operator='equal', formula=['1'], 
                  fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    
    # Add rows for historical quarters (user input)
    for i in range(5):
        row += 1
        ws[f"A{row}"] = f"[Enter Quarter]"
        for col in range(2, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Copy formulas for columns E and G
        ws[f"E{row}"] = f'=IF(C{row}=0,"N/A",D{row}/C{row})'
        ws[f"E{row}"].number_format = '0.0"%"'
        ws[f"G{row}"] = f'=IF(E{row}=1,"✅ Compliant",IF(E{row}>=0.9,"⚠️ Needs Attention","❌ Non-Compliant"))'
    
    row += 1
    
    # Target statement
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "🎯 TARGET: 100% of vulnerability-related risks must have CVSS quantification for Clause 6.1 compliance"
    ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True, color="0000FF")
    ws[f"A{row}"].fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30



# ============================================================================
# SECTION 4: SHEET 1 - EXECUTIVE SUMMARY (DASHBOARD)
# ============================================================================

def create_executive_summary(ws, styles):
    """
    Sheet 1: Executive Summary - Single-page dashboard for C-level.
    
    NOTE: This is a LAYOUT specification. Actual charts must be added manually
    or via separate charting libraries. We create the data structure and placeholders.
    """
    
    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = "Threat Intelligence Program - Executive Summary"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # Report period
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Reporting Period: {datetime.now().strftime('%B %Y')}"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    
    # Generation date
    ws.merge_cells("E2:H2")
    ws["E2"] = f"Generated: {datetime.now().strftime('%d.%m.%Y')}"
    ws["E2"].font = Font(name="Calibri", size=11, italic=True)
    ws["E2"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Classification
    ws.merge_cells("I2:L2")
    ws["I2"] = "Classification: INTERNAL"
    ws["I2"].font = Font(name="Calibri", size=11, italic=True, bold=True)
    ws["I2"].alignment = Alignment(horizontal="right", vertical="center")
    
    row = 4
    
    # Program Health Score (Large Metric Box)
    ws.merge_cells(f"A{row}:D{row+3}")
    ws[f"A{row}"] = "PROGRAM HEALTH SCORE"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    row += 4
    ws.merge_cells(f"A{row}:D{row+2}")
    # Composite score formula (weighted average from KPIs)
    ws[f"A{row}"] = '=AVERAGE(Program_KPIs!E5:E24)'
    ws[f"A{row}"].font = Font(name="Calibri", size=36, bold=True)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{row}"].number_format = '0.0'
    ws[f"A{row}"].border = get_border()
    
    row += 3
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Scale: 1.0 (Critical) → 5.0 (Excellent)"
    ws[f"A{row}"].font = Font(name="Calibri", size=9, italic=True)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Key Metrics Boxes (4 boxes across top)
    row = 4
    
    # Box 1: Intelligence Production
    ws.merge_cells(f"F{row}:H{row}")
    ws[f"F{row}"] = "Intelligence Production"
    ws[f"F{row}"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws[f"F{row}"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    ws.merge_cells(f"F{row}:H{row+1}")
    ws[f"F{row}"] = "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Intelligence_Production'!N:N,\"Published\")"
    ws[f"F{row}"].font = Font(name="Calibri", size=24, bold=True)
    ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"F{row}"].border = get_border()
    
    row += 2
    ws.merge_cells(f"F{row}:H{row}")
    ws[f"F{row}"] = "Products This Month"
    ws[f"F{row}"].font = Font(name="Calibri", size=9)
    ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Box 2: Stakeholder Satisfaction
    row = 4
    ws.merge_cells(f"I{row}:J{row}")
    ws[f"I{row}"] = "Stakeholder Satisfaction"
    ws[f"I{row}"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws[f"I{row}"].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    ws.merge_cells(f"I{row}:J{row+1}")
    ws[f"I{row}"] = "=AVERAGE('[ISMS-IMP-A.5.7.2.xlsx]Intelligence_Production'!J:J)"
    ws[f"I{row}"].font = Font(name="Calibri", size=24, bold=True)
    ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"I{row}"].number_format = '0.0'
    ws[f"I{row}"].border = get_border()
    
    row += 2
    ws.merge_cells(f"I{row}:J{row}")
    ws[f"I{row}"] = "Avg Rating (1-5)"
    ws[f"I{row}"].font = Font(name="Calibri", size=9)
    ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Box 3: IOC Effectiveness
    row = 4
    ws.merge_cells(f"K{row}:L{row}")
    ws[f"K{row}"] = "IOC Effectiveness"
    ws[f"K{row}"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws[f"K{row}"].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    ws.merge_cells(f"K{row}:L{row+1}")
    ws[f"K{row}"] = "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!A:A)-4"
    ws[f"K{row}"].font = Font(name="Calibri", size=24, bold=True)
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"K{row}"].border = get_border()
    
    row += 2
    ws.merge_cells(f"K{row}:L{row}")
    ws[f"K{row}"] = "IOCs Deployed"
    ws[f"K{row}"].font = Font(name="Calibri", size=9)
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Integration Status Section (row 14)
    row = 14
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "INTEGRATION STATUS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    ws[f"A{row}"] = "Security Tools Integrated:"
    ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"C{row}"] = "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!A:A)-4"
    ws[f"C{row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"C{row}"].border = get_border()
    
    row += 1
    ws[f"A{row}"] = "Fully Integrated:"
    ws[f"A{row}"].font = Font(name="Calibri", size=10)
    ws[f"C{row}"] = "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!D:D,\"Full\")"
    ws[f"C{row}"].font = Font(name="Calibri", size=10)
    ws[f"C{row}"].border = get_border()
    
    row += 1
    ws[f"A{row}"] = "Partially Integrated:"
    ws[f"A{row}"].font = Font(name="Calibri", size=10)
    ws[f"C{row}"] = "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!D:D,\"Partial\")"
    ws[f"C{row}"].font = Font(name="Calibri", size=10)
    ws[f"C{row}"].border = get_border()
    
    # Control 8.8 Linkage Section
    row = 14
    ws.merge_cells(f"F{row}:H{row}")
    ws[f"F{row}"] = "CONTROL 8.8 LINKAGE (VTL)"
    ws[f"F{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"F{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    ws[f"F{row}"] = "Total VT Links:"
    ws[f"F{row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"H{row}"] = "=COUNTA('[ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Linked_Threats'!A:A)-4"
    ws[f"H{row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"H{row}"].border = get_border()
    
    row += 1
    ws[f"F{row}"] = "Active Exploitation:"
    ws[f"F{row}"].font = Font(name="Calibri", size=10, color="FF0000", bold=True)
    ws[f"H{row}"] = "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Linked_Threats'!E:E,\"Active_Exploitation\")"
    ws[f"H{row}"].font = Font(name="Calibri", size=10, color="FF0000", bold=True)
    ws[f"H{row}"].border = get_border()
    
    row += 1
    ws[f"F{row}"] = "Patched:"
    ws[f"F{row}"].font = Font(name="Calibri", size=10)
    ws[f"H{row}"] = "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Linked_Threats'!R:R,\"Patched\")"
    ws[f"H{row}"].font = Font(name="Calibri", size=10)
    ws[f"H{row}"].border = get_border()
    
    # Compliance Status Section
    row = 14
    ws.merge_cells(f"I{row}:L{row}")
    ws[f"I{row}"] = "COMPLIANCE STATUS"
    ws[f"I{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"I{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    ws[f"I{row}"] = "ISO 27001:2022 Control 5.7:"
    ws[f"I{row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"L{row}"] = "=COUNTIF(Compliance_Evidence!F:F,\"Current\")/COUNTA(Compliance_Evidence!A:A)-4"
    ws[f"L{row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"L{row}"].number_format = '0%'
    ws[f"L{row}"].border = get_border()
    
    row += 1
    ws[f"I{row}"] = "Outstanding Evidence:"
    ws[f"I{row}"].font = Font(name="Calibri", size=10)
    ws[f"L{row}"] = "=COUNTIF(Compliance_Evidence!F:F,\"Missing\")"
    ws[f"L{row}"].font = Font(name="Calibri", size=10)
    ws[f"L{row}"].border = get_border()
    
    row += 1
    ws[f"I{row}"] = "Next Audit:"
    ws[f"I{row}"].font = Font(name="Calibri", size=10)
    ws[f"L{row}"] = "[Enter Date]"
    ws[f"L{row}"].font = Font(name="Calibri", size=10, italic=True)
    ws[f"L{row}"].fill = styles["input_cell"]["fill"]
    ws[f"L{row}"].border = get_border()
    
    # Chart placeholders (row 22)
    row = 22
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "📊 VISUALIZATION AREAS (Add charts manually or via reporting tools)"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    chart_areas = [
        "Area 1 (A23:D30): Source Performance Trend Chart (6 months)",
        "Area 2 (E23:H30): Intelligence Production Volume (bar chart)",
        "Area 3 (I23:L30): Stakeholder Satisfaction Trend (line chart)",
        "Area 4 (A31:F38): MITRE ATT&CK Coverage Heat Map",
        "Area 5 (G31:L38): Risk Indicator Dashboard (gauges)",
    ]
    
    for area in chart_areas:
        ws[f"A{row}"] = area
        ws[f"A{row}"].font = Font(name="Calibri", size=10, italic=True)
        row += 1
    
    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws.column_dimensions[col].width = 15
    
    # Note about external references
    row = 40
    ws.merge_cells(f"A{row}:L{row+3}")
    ws[f"A{row}"] = get_external_reference_note()
    ws[f"A{row}"].font = Font(name="Calibri", size=9, color="FF0000")
    ws[f"A{row}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 100


# ============================================================================
# END OF SECTION 1
# ============================================================================

# ============================================================================
# SECTION 2: SHEETS 2-4 (PROGRAM KPIs, SOURCE PORTFOLIO, INTELLIGENCE OPS)
# ============================================================================

def create_program_kpis(ws, styles, validations):
    """Sheet 2: Program KPIs - Key performance indicators with targets."""
    
    # Title
    ws.merge_cells("A1:M1")
    ws["A1"] = "Threat Intelligence Program - Key Performance Indicators"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:M2")
    ws["A2"] = "Track program KPIs with targets and thresholds. Set target values based on organizational goals."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        "KPI_ID", "KPI_Category", "KPI_Name", "Description", "Unit",
        "Measurement_Frequency", "Data_Source", "Target_Value", "Warning_Threshold",
        "Critical_Threshold", "Current_Value", "Last_Month_Value", "Performance_vs_Target",
        "Status", "Trend", "Owner", "Last_Updated", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 35
    
    # Pre-populate standard KPIs with external references
    kpis = [
        # Source Management (5.7.1)
        ("KPI-TI-001", "Source_Management", "Number of Active Sources", "Count of active TI sources", "Count", "Monthly",
         "=[ISMS-IMP-A.5.7.1.xlsx]Source_Inventory!B:B", "20", "16", "12"),
        
        ("KPI-TI-002", "Source_Management", "Average Source Quality Score", "Avg Admiralty Code score", "Score (1-5)", "Monthly",
         "=[ISMS-IMP-A.5.7.1.xlsx]Source_Evaluation!H:H", "4.0", "3.5", "3.0"),
        
        ("KPI-TI-003", "Source_Management", "Source Coverage Gaps", "Requirements with inadequate coverage", "Count", "Monthly",
         "=[ISMS-IMP-A.5.7.1.xlsx]Coverage_Matrix!Summary", "5", "8", "12"),
        
        ("KPI-TI-004", "Source_Management", "Cost Per Source (Avg)", "Average annual cost per source", "USD", "Quarterly",
         "=[ISMS-IMP-A.5.7.1.xlsx]Cost_Analysis!E:E", "50000", "60000", "75000"),
        
        # Collection & Analysis (5.7.2)
        ("KPI-TI-005", "Production", "Intelligence Products per Month", "Count of published products", "Reports", "Monthly",
         "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Intelligence_Production'!N:N,\"Published\")", "20", "16", "12"),
        
        ("KPI-TI-006", "Analysis", "Avg Time to Produce (hours)", "Hours from start to publication", "Hours", "Monthly",
         "=AVERAGE('[ISMS-IMP-A.5.7.2.xlsx]Intelligence_Production'!M:M)", "48", "72", "96"),
        
        ("KPI-TI-007", "Production", "Stakeholder Feedback Rating", "Average consumer rating", "Rating (1-5)", "Monthly",
         "=AVERAGE('[ISMS-IMP-A.5.7.2.xlsx]Intelligence_Production'!J:J)", "4.0", "3.5", "3.0"),
        
        ("KPI-TI-008", "Production", "Intelligence Consumption Rate", "% of products acted upon", "Percentage", "Monthly",
         "=AVERAGE('[ISMS-IMP-A.5.7.2.xlsx]Intelligence_Production'!K:K)", "75%", "60%", "50%"),
        
        ("KPI-TI-009", "Analysis", "MITRE ATT&CK Coverage", "% of critical techniques covered", "Percentage", "Quarterly",
         "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]MITRE_Mapping'!F:F,\"Full\")/COUNTA('[ISMS-IMP-A.5.7.2.xlsx]MITRE_Mapping'!A:A)-4", "80%", "70%", "60%"),
        
        ("KPI-TI-010", "Analysis", "Analyst Training Completion", "% of planned training completed", "Percentage", "Quarterly",
         "=AVERAGE('[ISMS-IMP-A.5.7.2.xlsx]Analyst_Capabilities'!J:J)", "90%", "75%", "60%"),
        
        ("KPI-TI-011", "Collection", "VT Links Created per Month", "Vulnerability-threat linkages", "Count", "Monthly",
         "=COUNTA('[ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Linked_Threats'!A:A)-4", "10", "7", "5"),
        
        # Integration & Distribution (5.7.3)
        ("KPI-TI-012", "Integration", "Security Tools Integrated", "Tools with TI integration", "Count", "Quarterly",
         "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!A:A)-4", "15", "12", "10"),
        
        ("KPI-TI-013", "Integration", "IOCs Deployed per Month", "Indicators deployed to security tools", "Count", "Monthly",
         "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!A:A)-4", "100", "75", "50"),
        
        ("KPI-TI-014", "Integration", "IOC Hit Rate", "% of IOCs with detections", "Percentage", "Monthly",
         "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!I:I,\"Active\")/COUNTA('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!A:A)-4", "15%", "10%", "5%"),
        
        ("KPI-TI-015", "Integration", "IOC False Positive Rate", "% of IOCs generating false positives", "Percentage", "Monthly",
         "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!J:J,\">0\")/COUNTA('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!A:A)-4", "5%", "10%", "15%"),
        
        ("KPI-TI-016", "Dissemination", "Stakeholder Engagement Score", "Average engagement rating", "Rating (1-5)", "Monthly",
         "=AVERAGE('[ISMS-IMP-A.5.7.3.xlsx]Feedback_Collection'!D:D)", "4.0", "3.5", "3.0"),
        
        ("KPI-TI-017", "Dissemination", "Distribution Reach", "% of target audience reached", "Percentage", "Monthly",
         "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Distribution_Tracking'!A:A)/COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Stakeholder_Registry'!A:A)", "90%", "75%", "60%"),
        
        ("KPI-TI-018", "Dissemination", "Mean Time to Action", "Hours from intel to action taken", "Hours", "Monthly",
         "=AVERAGE('[ISMS-IMP-A.5.7.3.xlsx]Distribution_Tracking'!TimeToAction)", "24", "48", "72"),
    ]
    
    # Add CVSS KPIs (NEW IN V1.0)
    cvss_kpis_list = get_cvss_kpis()
    kpis = list(kpis) + cvss_kpis_list
    
    
    row = 5
    for kpi_id, category, name, desc, unit, freq, data_source, target, warning, critical in kpis:
        ws[f"A{row}"] = kpi_id
        ws[f"B{row}"] = category
        ws[f"C{row}"] = name
        ws[f"D{row}"] = desc
        ws[f"E{row}"] = unit
        ws[f"F{row}"] = freq
        ws[f"G{row}"] = data_source  # External reference formula
        ws[f"H{row}"] = target
        ws[f"I{row}"] = warning
        ws[f"J{row}"] = critical
        
        # Format cells
        for col in range(1, 19):
            cell = ws.cell(row=row, column=col)
            if col in [1, 7, 13, 14, 15]:  # ID, Data_Source, Performance, Status, Trend - formulas
                cell.fill = styles["formula_cell"]["fill"]
            else:
                cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Current_Value (user input)
        ws[f"K{row}"].fill = styles["input_cell"]["fill"]
        
        # Last_Month_Value (user input)
        ws[f"L{row}"].fill = styles["input_cell"]["fill"]
        
        # Performance_vs_Target (formula)
        ws[f"M{row}"] = f'=IF(OR(K{row}="",H{row}=""),"N/A",(K{row}/H{row}-1)*100)'
        ws[f"M{row}"].number_format = '0.0"%"'
        
        # Status (formula based on thresholds)
        ws[f"N{row}"] = f'=IF(A{row}="KPI-CVSS-003",IF(K{row}="","N/A",IF(K{row}=0,"On_Track","At_Risk")),IF(K{row}="","N/A",IF(K{row}>=H{row},"On_Track",IF(K{row}>=I{row},"Caution","At_Risk"))))' # Special handling for emergency indicator
        # Fallback for non-CVSS-003 rows (this gets overwritten)
        if ws[f"A{row}"].value != "KPI-CVSS-003":
            ws[f"N{row}"] = f'=IF(K{row}="","N/A",IF(K{row}>=H{row},"On_Track",IF(K{row}>=I{row},"Caution","At_Risk")))'
        
        # Trend (formula - simple comparison)
        ws[f"O{row}"] = f'=IF(OR(K{row}="",L{row}=""),"N/A",IF(K{row}>L{row},"Improving",IF(K{row}<L{row},"Declining","Stable")))'
        
        # Owner, Last_Updated, Notes (user input)
        for col in [16, 17, 18]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        ws[f"Q{row}"].number_format = 'YYYY-MM-DD'
        
        row += 1
    
    # Additional blank rows for custom KPIs (rows 23-30)
    for row in range(row, row+8):
        ws[f"A{row}"] = f"KPI-TI-{str(row-4).zfill(3)}"
        
        for col in range(1, 19):
            cell = ws.cell(row=row, column=col)
            if col in [1, 13, 14, 15]:
                cell.fill = styles["formula_cell"]["fill"]
                if col == 13:
                    cell.value = f'=IF(OR(K{row}="",H{row}=""),"N/A",(K{row}/H{row}-1)*100)'
                    cell.number_format = '0.0"%"'
                elif col == 14:
                    cell.value = f'=IF(K{row}="","N/A",IF(K{row}>=H{row},"On_Track",IF(K{row}>=I{row},"Caution","At_Risk")))'
                elif col == 15:
                    cell.value = f'=IF(OR(K{row}="",L{row}=""),"N/A",IF(K{row}>L{row},"Improving",IF(K{row}<L{row},"Declining","Stable")))'
            else:
                cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
    
    # Apply validations
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    # Apply validations to specific columns
    for row in range(5, 31):
        validations['kpi_category'].add(ws[f"B{row}"])
        validations['measurement_frequency'].add(ws[f"F{row}"])
    
    # Conditional formatting for Status
    ws.conditional_formatting.add(
        'N5:N40',
        CellIsRule(operator='equal', formula=['"On_Track"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    
    # Special RED highlighting for KPI-CVSS-003 emergency status (NEW IN V1.0)
    for check_row in range(5, 40):
        if ws[f"A{check_row}"].value == "KPI-CVSS-003":
            ws.conditional_formatting.add(
                f'K{check_row}:N{check_row}',
                CellIsRule(operator='greaterThan', formula=['0'],
                          fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                          font=Font(color="FFFFFF", bold=True))
            )
            break
    
    ws.conditional_formatting.add(
        'N5:N40',
        CellIsRule(operator='equal', formula=['"Caution"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'N5:N40',
        CellIsRule(operator='equal', formula=['"At_Risk"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    
    # Summary dashboard (row 32)
    ws.merge_cells("A32:D32")
    ws["A32"] = "KPI PERFORMANCE SUMMARY"
    ws["A32"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A32"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    summary = [
        ("Total KPIs Tracked:", '=COUNTA(A5:A30)'),
        ("On Track:", '=COUNTIF(N5:N30,"On_Track")'),
        ("Caution:", '=COUNTIF(N5:N30,"Caution")'),
        ("At Risk:", '=COUNTIF(N5:N30,"At_Risk")'),
        ("Improving Trends:", '=COUNTIF(O5:O30,"Improving")'),
        ("Declining Trends:", '=COUNTIF(O5:O30,"Declining")'),
    ]
    
    row = 33
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 18
    ws.column_dimensions['N'].width = 12
    ws.column_dimensions['O'].width = 12
    ws.column_dimensions['P'].width = 25
    ws.column_dimensions['Q'].width = 14
    ws.column_dimensions['R'].width = 30
    
    ws.freeze_panes = "A5"


def create_source_portfolio(ws, styles):
    """Sheet 3: Source Portfolio - Aggregated from 5.7.1."""
    
    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "Source Portfolio Performance (from ISMS-IMP-A.5.7.1)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:J2")
    ws["A2"] = "Read-only aggregation from 5.7.1 Sources workbook. Update source workbook to refresh data here."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="0000FF")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # External reference note
    ws.merge_cells("A3:J3")
    ws["A3"] = "⚠️ External Reference: [ISMS-IMP-A.5.7.1.xlsx]"
    ws["A3"].font = Font(name="Calibri", size=10, bold=True, color="FF0000")
    ws["A3"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Summary metrics (row 5)
    ws.merge_cells("A5:J5")
    ws["A5"] = "SOURCE PORTFOLIO SUMMARY"
    ws["A5"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A5"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    metrics = [
        ("Total Active Sources:", "=COUNTA('[ISMS-IMP-A.5.7.1.xlsx]Source_Inventory'!A:A)-4"),
        ("Commercial Sources:", "=COUNTIF('[ISMS-IMP-A.5.7.1.xlsx]Source_Inventory'!C:C,\"Commercial\")"),
        ("OSINT Sources:", "=COUNTIF('[ISMS-IMP-A.5.7.1.xlsx]Source_Inventory'!C:C,\"OSINT\")"),
        ("Government Sources:", "=COUNTIF('[ISMS-IMP-A.5.7.1.xlsx]Source_Inventory'!C:C,\"Government\")"),
        ("Average Quality Score:", "=AVERAGE('[ISMS-IMP-A.5.7.1.xlsx]Source_Evaluation'!H:H)"),
        ("Sources Rated Excellent:", "=COUNTIF('[ISMS-IMP-A.5.7.1.xlsx]Source_Evaluation'!L:L,\"Excellent\")"),
        ("Coverage Gaps Identified:", "=COUNTIF('[ISMS-IMP-A.5.7.1.xlsx]Coverage_Matrix'!Status,\"Gap\")"),
        ("Annual Portfolio Cost:", "=SUM('[ISMS-IMP-A.5.7.1.xlsx]Cost_Analysis'!E:E)"),
        ("Compliance Issues:", "=COUNTIF('[ISMS-IMP-A.5.7.1.xlsx]Compliance_Check'!I:I,\"Non_Compliant\")"),
    ]
    
    row = 6
    for label, formula in metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"C{row}"] = formula
        ws[f"C{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"C{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"C{row}"].border = get_border()
        
        if "Cost" in label:
            ws[f"C{row}"].number_format = '#,##0.00'
        elif "Quality" in label:
            ws[f"C{row}"].number_format = '0.0'
        
        row += 1
    
    # Top Sources Table (row 17)
    ws.merge_cells("A17:J17")
    ws["A17"] = "TOP PERFORMING SOURCES (Quality Score)"
    ws["A17"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A17"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Column headers
    headers = ["Source_ID", "Source_Name", "Type", "Quality_Score", "Coverage_Rating", 
               "Cost_Annual", "ROI_Rating", "Recommendation", "Status", "Notes"]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=18, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    # Note: Actual data would require more complex external formulas or manual entry
    # For demonstration, we show the structure
    ws.merge_cells("A19:J19")
    ws["A19"] = "Note: Top sources require manual aggregation or advanced Excel formulas (SORT, FILTER - Excel 365)"
    ws["A19"].font = Font(name="Calibri", size=9, italic=True)
    ws["A19"].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 30
    
    ws.freeze_panes = "A19"


def create_intelligence_operations(ws, styles):
    """Sheet 4: Intelligence Operations - Aggregated from 5.7.2."""
    
    # Title
    ws.merge_cells("A1:K1")
    ws["A1"] = "Intelligence Operations Metrics (from ISMS-IMP-A.5.7.2)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:K2")
    ws["A2"] = "Read-only aggregation from 5.7.2 Collection & Analysis workbook. Includes VTL metrics."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="0000FF")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # External reference note
    ws.merge_cells("A3:K3")
    ws["A3"] = "⚠️ External Reference: [ISMS-IMP-A.5.7.2.xlsx]"
    ws["A3"].font = Font(name="Calibri", size=10, bold=True, color="FF0000")
    ws["A3"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Collection Coverage Summary
    ws.merge_cells("A5:E5")
    ws["A5"] = "COLLECTION COVERAGE"
    ws["A5"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A5"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    coverage_metrics = [
        ("Total Requirements:", "=COUNTA('[ISMS-IMP-A.5.7.2.xlsx]Collection_Coverage'!A:A)-4"),
        ("Adequate Coverage:", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Collection_Coverage'!L:L,\"Adequate\")"),
        ("Minimal Coverage:", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Collection_Coverage'!L:L,\"Minimal\")"),
        ("Coverage Gaps:", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Collection_Coverage'!L:L,\"Gap\")"),
    ]
    
    row = 6
    for label, formula in coverage_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"C{row}"] = formula
        ws[f"C{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"C{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"C{row}"].border = get_border()
        row += 1
    
    # Intelligence Production Summary
    ws.merge_cells("G5:K5")
    ws["G5"] = "INTELLIGENCE PRODUCTION"
    ws["G5"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["G5"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    production_metrics = [
        ("Total Products:", "=COUNTA('[ISMS-IMP-A.5.7.2.xlsx]Intelligence_Production'!A:A)-4"),
        ("Published This Month:", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Intelligence_Production'!N:N,\"Published\")"),
        ("Avg Feedback Rating:", "=AVERAGE('[ISMS-IMP-A.5.7.2.xlsx]Intelligence_Production'!J:J)"),
        ("Avg Time to Produce:", "=AVERAGE('[ISMS-IMP-A.5.7.2.xlsx]Intelligence_Production'!M:M)"),
    ]
    
    row = 6
    for label, formula in production_metrics:
        ws[f"G{row}"] = label
        ws[f"G{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"I{row}"] = formula
        ws[f"I{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"I{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"I{row}"].border = get_border()
        
        if "Rating" in label:
            ws[f"I{row}"].number_format = '0.0'
        elif "Time" in label:
            ws[f"I{row}"].number_format = '0.0" hrs"'
        
        row += 1
    
    # MITRE ATT&CK Coverage
    ws.merge_cells("A12:E12")
    ws["A12"] = "MITRE ATT&CK COVERAGE"
    ws["A12"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A12"].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    
    mitre_metrics = [
        ("Total Techniques Tracked:", "=COUNTA('[ISMS-IMP-A.5.7.2.xlsx]MITRE_Mapping'!A:A)-4"),
        ("Full Coverage:", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]MITRE_Mapping'!F:F,\"Full\")"),
        ("Partial Coverage:", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]MITRE_Mapping'!F:F,\"Partial\")"),
        ("No Coverage:", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]MITRE_Mapping'!F:F,\"None\")"),
    ]
    
    row = 13
    for label, formula in mitre_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"C{row}"] = formula
        ws[f"C{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"C{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"C{row}"].border = get_border()
        row += 1
    
    # VTL Summary (CRITICAL)
    ws.merge_cells("G12:K12")
    ws["G12"] = "VULNERABILITY-THREAT LINKS (VTL) ⚠️"
    ws["G12"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["G12"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    
    vtl_metrics = [
        ("Total VT Links:", "=COUNTA('[ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Linked_Threats'!A:A)-4"),
        ("Active Exploitation:", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Linked_Threats'!E:E,\"Active_Exploitation\")"),
        ("Critical Assets Affected:", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Linked_Threats'!Q:Q,\"Yes\")"),
        ("Remediation - Open:", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Linked_Threats'!R:R,\"Open\")"),
        ("Remediation - Patched:", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Linked_Threats'!R:R,\"Patched\")"),
    ]
    
    row = 13
    for label, formula in vtl_metrics:
        ws[f"G{row}"] = label
        ws[f"G{row}"].font = Font(name="Calibri", size=10, bold=True)
        if "Active" in label or "Critical" in label:
            ws[f"G{row}"].font = Font(name="Calibri", size=10, bold=True, color="FF0000")
        ws[f"I{row}"] = formula
        ws[f"I{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"I{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"I{row}"].border = get_border()
        row += 1
    
    # Analyst Team Summary
    ws.merge_cells("A19:K19")
    ws["A19"] = "ANALYST TEAM CAPABILITY"
    ws["A19"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A19"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    team_metrics = [
        ("Total Analysts:", "=COUNTA('[ISMS-IMP-A.5.7.2.xlsx]Analyst_Capabilities'!A:A)-4"),
        ("Avg Years TI Experience:", "=AVERAGE('[ISMS-IMP-A.5.7.2.xlsx]Analyst_Capabilities'!F:F)"),
        ("Training Completion Rate:", "=AVERAGE('[ISMS-IMP-A.5.7.2.xlsx]Analyst_Capabilities'!J:J)"),
        ("Certifications Held:", "=COUNTA('[ISMS-IMP-A.5.7.2.xlsx]Analyst_Capabilities'!Q:Q)-4"),
    ]
    
    row = 20
    for label, formula in team_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"D{row}"] = formula
        ws[f"D{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"D{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"D{row}"].border = get_border()
        
        if "Years" in label:
            ws[f"D{row}"].number_format = '0.0'
        elif "Rate" in label:
            ws[f"D{row}"].number_format = '0%'
        
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15


# ============================================================================
# END OF SECTION 2
# ============================================================================

# ============================================================================
# SECTION 3: SHEETS 5-8 (INTEGRATION, STAKEHOLDER, TRENDS, RISKS)
# ============================================================================

def create_integration_status(ws, styles):
    """Sheet 5: Integration Status - Aggregated from 5.7.3."""
    
    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "Integration Status (from ISMS-IMP-A.5.7.3)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:J2")
    ws["A2"] = "Read-only aggregation from 5.7.3 Integration & Distribution workbook. Tool integration and IOC deployment metrics."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="0000FF")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # External reference note
    ws.merge_cells("A3:J3")
    ws["A3"] = "⚠️ External Reference: [ISMS-IMP-A.5.7.3.xlsx]"
    ws["A3"].font = Font(name="Calibri", size=10, bold=True, color="FF0000")
    ws["A3"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Tool Integration Summary
    ws.merge_cells("A5:E5")
    ws["A5"] = "SECURITY TOOL INTEGRATION"
    ws["A5"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A5"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    tool_metrics = [
        ("Total Security Tools:", "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!A:A)-4"),
        ("Fully Integrated:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!D:D,\"Full\")"),
        ("Partially Integrated:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!D:D,\"Partial\")"),
        ("Planned Integration:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!D:D,\"Planned\")"),
        ("Not Integrated:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!D:D,\"None\")"),
        ("Integration Coverage %:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!D:D,\"Full\")/COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!A:A)-4"),
    ]
    
    row = 6
    for label, formula in tool_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"C{row}"] = formula
        ws[f"C{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"C{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"C{row}"].border = get_border()
        
        if "%" in label:
            ws[f"C{row}"].number_format = '0%'
        
        row += 1
    
    # IOC Deployment Summary
    ws.merge_cells("F5:J5")
    ws["F5"] = "IOC DEPLOYMENT & EFFECTIVENESS"
    ws["F5"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["F5"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    ioc_metrics = [
        ("Total IOCs Deployed:", "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!A:A)-4"),
        ("Active IOCs:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!I:I,\"Active\")"),
        ("Expired IOCs:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!I:I,\"Expired\")"),
        ("IOCs with Hits:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!J:J,\">0\")"),
        ("Hit Rate %:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!J:J,\">0\")/COUNTA('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!A:A)-4"),
        ("Avg False Positives:", "=AVERAGE('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!K:K)"),
    ]
    
    row = 6
    for label, formula in ioc_metrics:
        ws[f"F{row}"] = label
        ws[f"F{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"H{row}"] = formula
        ws[f"H{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"H{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"H{row}"].border = get_border()
        
        if "%" in label:
            ws[f"H{row}"].number_format = '0%'
        elif "Avg" in label:
            ws[f"H{row}"].number_format = '0.0'
        
        row += 1
    
    # Dissemination Channels Summary
    ws.merge_cells("A14:J14")
    ws["A14"] = "DISSEMINATION CHANNELS"
    ws["A14"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A14"].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    
    channel_metrics = [
        ("Total Channels Configured:", "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Dissemination_Channels'!A:A)-4"),
        ("Active Channels:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Dissemination_Channels'!G:G,\"Active\")"),
        ("Distribution Events (Month):", "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Distribution_Tracking'!A:A)-4"),
        ("Avg Distribution Time (hrs):", "=AVERAGE('[ISMS-IMP-A.5.7.3.xlsx]Distribution_Tracking'!TimeCol)"),
    ]
    
    row = 15
    for label, formula in channel_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"D{row}"] = formula
        ws[f"D{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"D{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"D{row}"].border = get_border()
        
        if "hrs" in label:
            ws[f"D{row}"].number_format = '0.0" hrs"'
        
        row += 1
    
    # Integration Health Dashboard
    ws.merge_cells("A21:J21")
    ws["A21"] = "INTEGRATION HEALTH INDICATORS"
    ws["A21"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A21"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    health_metrics = [
        ("Tools Requiring Attention:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!J:J,\"High\")"),
        ("Integration Issues Open:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!Issues,\">0\")"),
        ("IOCs Pending Deployment:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]IOC_Deployment'!I:I,\"Pending\")"),
        ("Average Integration Maturity:", "=AVERAGE(IntegrationMaturityScores)"),
    ]
    
    row = 22
    for label, formula in health_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"D{row}"] = formula if "COUNTIF" in formula else "Manual Calculation Required"
        ws[f"D{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"D{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"D{row}"].border = get_border()
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15


def create_stakeholder_engagement(ws, styles):
    """Sheet 6: Stakeholder Engagement - Aggregated from 5.7.3."""
    
    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "Stakeholder Engagement Metrics (from ISMS-IMP-A.5.7.3)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:H2")
    ws["A2"] = "Read-only aggregation from 5.7.3 Integration & Distribution workbook. Stakeholder feedback and engagement tracking."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="0000FF")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # External reference note
    ws.merge_cells("A3:H3")
    ws["A3"] = "⚠️ External Reference: [ISMS-IMP-A.5.7.3.xlsx]"
    ws["A3"].font = Font(name="Calibri", size=10, bold=True, color="FF0000")
    ws["A3"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Stakeholder Registry Summary
    ws.merge_cells("A5:H5")
    ws["A5"] = "STAKEHOLDER REGISTRY"
    ws["A5"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A5"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    registry_metrics = [
        ("Total Registered Stakeholders:", "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Stakeholder_Registry'!A:A)-4"),
        ("Executive Level:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Stakeholder_Registry'!D:D,\"Executive\")"),
        ("Management Level:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Stakeholder_Registry'!D:D,\"Management\")"),
        ("Technical Level:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Stakeholder_Registry'!D:D,\"Technical\")"),
        ("Active Stakeholders:", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Stakeholder_Registry'!Status,\"Active\")"),
    ]
    
    row = 6
    for label, formula in registry_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"C{row}"] = formula
        ws[f"C{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"C{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"C{row}"].border = get_border()
        row += 1
    
    # Feedback Collection Summary
    ws.merge_cells("A13:H13")
    ws["A13"] = "FEEDBACK & SATISFACTION"
    ws["A13"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A13"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    feedback_metrics = [
        ("Total Feedback Entries:", "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Feedback_Collection'!A:A)-4"),
        ("Avg Relevance Rating (1-5):", "=AVERAGE('[ISMS-IMP-A.5.7.3.xlsx]Feedback_Collection'!D:D)"),
        ("Avg Timeliness Rating (1-5):", "=AVERAGE('[ISMS-IMP-A.5.7.3.xlsx]Feedback_Collection'!E:E)"),
        ("Avg Actionability Rating (1-5):", "=AVERAGE('[ISMS-IMP-A.5.7.3.xlsx]Feedback_Collection'!F:F)"),
        ("Overall Satisfaction (Avg):", "=AVERAGE('[ISMS-IMP-A.5.7.3.xlsx]Feedback_Collection'!D:F)"),
        ("Feedback Response Rate %:", "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Feedback_Collection'!A:A)/COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Distribution_Tracking'!A:A)"),
    ]
    
    row = 14
    for label, formula in feedback_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"D{row}"] = formula
        ws[f"D{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"D{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"D{row}"].border = get_border()
        
        if "Rating" in label or "Satisfaction" in label:
            ws[f"D{row}"].number_format = '0.0'
        elif "%" in label:
            ws[f"D{row}"].number_format = '0%'
        
        row += 1
    
    # Engagement Trends
    ws.merge_cells("A22:H22")
    ws["A22"] = "ENGAGEMENT INDICATORS"
    ws["A22"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A22"].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    
    engagement_metrics = [
        ("Intelligence Consumption Rate:", "=AVERAGE('[ISMS-IMP-A.5.7.2.xlsx]Intelligence_Production'!K:K)"),
        ("Stakeholder Reach This Month:", "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Distribution_Tracking'!UniqueStakeholders)"),
        ("Distribution Events This Month:", "=COUNTA('[ISMS-IMP-A.5.7.3.xlsx]Distribution_Tracking'!A:A)-4"),
        ("Avg Time to Stakeholder Action:", "=AVERAGE('[ISMS-IMP-A.5.7.3.xlsx]Distribution_Tracking'!ActionTime)"),
    ]
    
    row = 23
    for label, formula in engagement_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"D{row}"] = formula if "AVERAGE" in formula or "COUNT" in formula else "Manual Calculation Required"
        ws[f"D{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"D{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"D{row}"].border = get_border()
        
        if "%" in label or "Rate" in label:
            ws[f"D{row}"].number_format = '0%'
        elif "Time" in label:
            ws[f"D{row}"].number_format = '0.0" hrs"'
        
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15


def create_trend_analysis(ws, styles):
    """Sheet 7: Trend Analysis - Historical tracking and forecasting."""
    
    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "Trend Analysis & Forecasting"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:H2")
    ws["A2"] = "Track KPI trends over time. Manually update monthly data for trend analysis. Historical data enables forecasting."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Monthly Trend Data Table (last 12 months)
    ws.merge_cells("A4:H4")
    ws["A4"] = "MONTHLY TREND DATA (Last 12 Months)"
    ws["A4"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Column headers
    headers = [
        "Month", "Intel_Products", "Avg_Feedback_Rating", "IOCs_Deployed", 
        "IOC_Hit_Rate_%", "VT_Links_Created", "Active_Sources", "Program_Health_Score"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    # Generate 12 month rows (starting from 12 months ago)
    current_date = datetime.now()
    for i in range(12, 0, -1):
        row = 6 + (12 - i)
        month_date = current_date - timedelta(days=30*i)
        
        # Month
        ws[f"A{row}"] = month_date.strftime("%Y-%m")
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = get_border()
        
        # Data cells (user input for historical data)
        for col in range(2, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            
            if col == 5:  # IOC_Hit_Rate
                cell.number_format = '0%'
            elif col == 3:  # Avg_Feedback_Rating
                cell.number_format = '0.0'
            elif col == 8:  # Program_Health_Score
                cell.number_format = '0.0'
    
    # Trend Summary (row 19)
    ws.merge_cells("A19:H19")
    ws["A19"] = "TREND SUMMARY"
    ws["A19"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A19"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    trend_metrics = [
        ("Avg Intel Products (12mo):", "=AVERAGE(B6:B17)"),
        ("Trend - Intel Products:", "=IF(B17>B6,\"Increasing\",IF(B17<B6,\"Decreasing\",\"Stable\"))"),
        ("Avg Feedback Rating (12mo):", "=AVERAGE(C6:C17)"),
        ("Trend - Feedback:", "=IF(C17>C6,\"Improving\",IF(C17<C6,\"Declining\",\"Stable\"))"),
        ("Avg IOC Hit Rate (12mo):", "=AVERAGE(E6:E17)"),
        ("Avg Program Health (12mo):", "=AVERAGE(H6:H17)"),
    ]
    
    row = 20
    for label, formula in trend_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"C{row}"] = formula
        ws[f"C{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"C{row}"].fill = styles["formula_cell"]["fill"]
        ws[f"C{row}"].border = get_border()
        
        if "Rating" in label or "Health" in label:
            ws[f"C{row}"].number_format = '0.0'
        elif "Rate" in label:
            ws[f"C{row}"].number_format = '0%'
        
        row += 1
    
    # Forecasting Note
    ws.merge_cells("A28:H28")
    ws["A28"] = "📈 FORECASTING & TREND CHARTS"
    ws["A28"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A28"].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    
    ws.merge_cells("A29:H31")
    forecast_note = """Trend charts should be created manually using the data above:
- Line chart: Intel Products over time (B6:B17)
- Line chart: Feedback Rating trend (C6:C17)
- Line chart: Program Health Score (H6:H17)

Use Excel's FORECAST.LINEAR function or trendlines for predictions."""
    ws["A29"] = forecast_note
    ws["A29"].font = Font(name="Calibri", size=10, italic=True)
    ws["A29"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[29].height = 60
    
    # Set column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 18
    
    ws.freeze_panes = "A6"


def create_risk_indicators(ws, styles, validations):
    """Sheet 8: Risk Indicators - Early warning for program degradation."""
    
    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "Risk Indicators & Early Warnings"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:I2")
    ws["A2"] = "Monitor risk indicators for TI program health. Set thresholds to trigger early warnings."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        "Risk_ID", "Risk_Category", "Risk_Indicator", "Current_Value",
        "Warning_Threshold", "Critical_Threshold", "Risk_Status",
        "Mitigation_Action", "Owner"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 30
    
    # Pre-populate standard risk indicators
    risks = [
        ("RISK-001", "Coverage", "Source Coverage Gaps", "=COUNTIF('[ISMS-IMP-A.5.7.1.xlsx]Coverage_Matrix'!Status,\"Gap\")", "5", "10"),
        ("RISK-002", "Production", "Intel Products Below Target", "=Program_KPIs!K5", "16", "12"),
        ("RISK-003", "Quality", "Stakeholder Satisfaction Declining", "=Program_KPIs!K7", "3.5", "3.0"),
        ("RISK-004", "Integration", "IOC Hit Rate Below Target", "=Program_KPIs!K14", "10%", "5%"),
        ("RISK-005", "Integration", "IOC False Positive Rate High", "=Program_KPIs!K15", "10%", "15%"),
        ("RISK-006", "Vulnerability", "Active Exploits Unpatched", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Linked_Threats'!R:R,\"Open\")", "3", "5"),
        ("RISK-007", "Team", "Analyst Training Incomplete", "=Program_KPIs!K10", "75%", "60%"),
        ("RISK-008", "Coverage", "MITRE Coverage Gaps", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]MITRE_Mapping'!F:F,\"None\")", "20", "30"),
        ("RISK-009", "Compliance", "Evidence Items Outdated", "=COUNTIF(Compliance_Evidence!F:F,\"Outdated\")", "3", "5"),
        ("RISK-010", "Integration", "Tools Not Integrated", "=COUNTIF('[ISMS-IMP-A.5.7.3.xlsx]Tool_Integration_Matrix'!D:D,\"None\")", "5", "8"),
    ]
    
    row = 5
    for risk_id, category, indicator, current_formula, warning, critical in risks:
        ws[f"A{row}"] = risk_id
        ws[f"B{row}"] = category
        ws[f"C{row}"] = indicator
        ws[f"D{row}"] = current_formula
        ws[f"E{row}"] = warning
        ws[f"F{row}"] = critical
        
        # Risk_Status formula
        ws[f"G{row}"] = f'=IF(D{row}="","N/A",IF(D{row}>F{row},"Critical",IF(D{row}>E{row},"Warning","Normal")))'
        
        # Format cells
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            if col in [1, 4, 7]:  # Risk_ID, Current_Value, Risk_Status
                cell.fill = styles["formula_cell"]["fill"]
            else:
                cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        row += 1
    
    # Additional rows for custom risks (rows 15-20)
    for row in range(15, 21):
        ws[f"A{row}"] = f"RISK-{str(row-4).zfill(3)}"
        ws[f"G{row}"] = f'=IF(D{row}="","N/A",IF(D{row}>F{row},"Critical",IF(D{row}>E{row},"Warning","Normal")))'
        
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            if col in [1, 4, 7]:
                cell.fill = styles["formula_cell"]["fill"]
            else:
                cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
    
    # Apply validations
    for row in range(5, 21):
        validations['risk_status'].add(ws[f"G{row}"])
    
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    # Conditional formatting for Risk_Status
    ws.conditional_formatting.add(
        'G5:G20',
        CellIsRule(operator='equal', formula=['"Normal"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'G5:G20',
        CellIsRule(operator='equal', formula=['"Warning"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'G5:G20',
        CellIsRule(operator='equal', formula=['"Critical"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    
    # Summary dashboard (row 22)
    ws.merge_cells("A22:D22")
    ws["A22"] = "RISK SUMMARY"
    ws["A22"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A22"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    
    summary = [
        ("Total Risks Monitored:", '=COUNTA(A5:A20)'),
        ("Critical Status:", '=COUNTIF(G5:G20,"Critical")'),
        ("Warning Status:", '=COUNTIF(G5:G20,"Warning")'),
        ("Normal Status:", '=COUNTIF(G5:G20,"Normal")'),
    ]
    
    row = 23
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        if "Critical" in label:
            ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True, color="FF0000")
            ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True, color="FF0000")
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 35
    ws.column_dimensions['I'].width = 25
    
    ws.freeze_panes = "A5"


# ============================================================================
# END OF SECTION 3
# ============================================================================

# ============================================================================
# SECTION 4: SHEETS 9-12 (COMPLIANCE, REPORTS, METADATA) + MAIN
# ============================================================================


    
    # Add CVSS quantification tracking (NEW IN V1.0)
    add_cvss_quantification_tracking(ws, styles, start_row=30)


def create_compliance_evidence(ws, styles, validations):
    """Sheet 9: Compliance Evidence - Audit artifacts for ISO 27001."""
    
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "Compliance Evidence - ISO 27001:2022 Control 5.7"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:G2")
    ws["A2"] = "Audit evidence inventory for ISO 27001:2022 Control 5.7 (Threat Intelligence). Track evidence status and currency."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        "Evidence_ID", "Control_Requirement", "Evidence_Type", "Location",
        "Last_Updated", "Status", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 30
    
    # Pre-populate standard evidence items
    evidence_items = [
        ("EV-5.7-001", "Source inventory maintained", "Excel Workbook", "ISMS-IMP-A.5.7.1 Sheet 2"),
        ("EV-5.7-002", "Source quality assessed", "Excel Workbook", "ISMS-IMP-A.5.7.1 Sheet 3"),
        ("EV-5.7-003", "Coverage gaps identified", "Excel Workbook", "ISMS-IMP-A.5.7.1 Sheet 4"),
        ("EV-5.7-004", "Cost-benefit analysis performed", "Excel Workbook", "ISMS-IMP-A.5.7.1 Sheet 5"),
        ("EV-5.7-005", "Compliance verification documented", "Excel Workbook", "ISMS-IMP-A.5.7.1 Sheet 6"),
        ("EV-5.7-006", "Intelligence requirements defined", "Excel Workbook", "ISMS-IMP-A.5.7.2 Sheet 2"),
        ("EV-5.7-007", "Analysis frameworks documented", "Excel Workbook", "ISMS-IMP-A.5.7.2 Sheet 3"),
        ("EV-5.7-008", "Analyst capabilities tracked", "Excel Workbook", "ISMS-IMP-A.5.7.2 Sheet 4"),
        ("EV-5.7-009", "Intelligence production tracked", "Excel Workbook", "ISMS-IMP-A.5.7.2 Sheet 5"),
        ("EV-5.7-010", "MITRE mapping maintained", "Excel Workbook", "ISMS-IMP-A.5.7.2 Sheet 6"),
        ("EV-5.7-011", "Quality metrics established", "Excel Workbook", "ISMS-IMP-A.5.7.2 Sheet 7"),
        ("EV-5.7-012", "Vulnerability linkage (VTL schema)", "Excel Workbook", "ISMS-IMP-A.5.7.2 Sheet 8 (CRITICAL)"),
        ("EV-5.7-013", "Process maturity assessed", "Excel Workbook", "ISMS-IMP-A.5.7.2 Sheet 9"),
        ("EV-5.7-014", "Tool integration documented", "Excel Workbook", "ISMS-IMP-A.5.7.3 Sheet 2"),
        ("EV-5.7-015", "IOC deployment tracked", "Excel Workbook", "ISMS-IMP-A.5.7.3 Sheet 3"),
        ("EV-5.7-016", "Dissemination channels configured", "Excel Workbook", "ISMS-IMP-A.5.7.3 Sheet 4"),
        ("EV-5.7-017", "Stakeholder engagement measured", "Excel Workbook", "ISMS-IMP-A.5.7.3 Sheet 7"),
        ("EV-5.7-018", "Program KPIs tracked", "Excel Workbook", "ISMS-IMP-A.5.7.4 Sheet 2"),
        ("EV-5.7-019", "Trend analysis performed", "Excel Workbook", "ISMS-IMP-A.5.7.4 Sheet 7"),
        ("EV-5.7-020", "Risk indicators monitored", "Excel Workbook", "ISMS-IMP-A.5.7.4 Sheet 8"),
    ]
    
    row = 5
    for ev_id, requirement, ev_type, location in evidence_items:
        ws[f"A{row}"] = ev_id
        ws[f"B{row}"] = requirement
        ws[f"C{row}"] = ev_type
        ws[f"D{row}"] = location
        
        # Format cells
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            if col == 1:  # Evidence_ID
                cell.fill = styles["formula_cell"]["fill"]
            else:
                cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Last_Updated (date)
        ws[f"E{row}"].number_format = 'YYYY-MM-DD'
        
        # Status dropdown
        validations['evidence_status'].add(ws[f"F{row}"])
        
        row += 1
    
    # Apply validations
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    # Conditional formatting for Status
    ws.conditional_formatting.add(
        'F5:F24',
        CellIsRule(operator='equal', formula=['"Current"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'F5:F24',
        CellIsRule(operator='equal', formula=['"Outdated"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'F5:F24',
        CellIsRule(operator='equal', formula=['"Missing"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    
    # Summary dashboard (row 26)
    ws.merge_cells("A26:D26")
    ws["A26"] = "COMPLIANCE SUMMARY"
    ws["A26"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A26"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    summary = [
        ("Total Evidence Items:", '=COUNTA(A5:A24)'),
        ("Current (< 90 days):", '=COUNTIF(F5:F24,"Current")'),
        ("Outdated (> 90 days):", '=COUNTIF(F5:F24,"Outdated")'),
        ("Missing:", '=COUNTIF(F5:F24,"Missing")'),
        ("Compliance Rate %:", '=COUNTIF(F5:F24,"Current")/COUNTA(A5:A24)'),
    ]
    
    row = 27
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        
        if "%" in label:
            ws[f"B{row}"].number_format = '0%'
        
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 40
    
    ws.freeze_panes = "A5"


def create_monthly_report(ws, styles):
    """Sheet 10: Monthly Report - Auto-generated monthly summary."""
    
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "Threat Intelligence Program - Monthly Report"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Report metadata
    ws.merge_cells("A2:C2")
    ws["A2"] = f"Report Period: {datetime.now().strftime('%B %Y')}"
    ws["A2"].font = Font(name="Calibri", size=11, bold=True)
    
    ws.merge_cells("D2:F2")
    ws["D2"] = f"Generated: {datetime.now().strftime('%d.%m.%Y')}"
    ws["D2"].font = Font(name="Calibri", size=11, italic=True)
    ws["D2"].alignment = Alignment(horizontal="right")
    
    row = 4
    
    # Section 1: Executive Summary
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "1. EXECUTIVE SUMMARY"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    row += 1
    ws.merge_cells(f"A{row}:F{row+2}")
    summary_text = """Program Health Score: [Auto from Executive_Summary]
Key Achievements: [User completes]
Areas Requiring Attention: [From Risk_Indicators]"""
    ws[f"A{row}"] = summary_text
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws[f"A{row}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws.row_dimensions[row].height = 60
    
    row += 3
    
    # Section 2: Key Metrics
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "2. KEY PERFORMANCE METRICS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    row += 1
    metrics = [
        ("Intelligence Products Published:", "=Program_KPIs!K5"),
        ("Stakeholder Satisfaction:", "=Program_KPIs!K7"),
        ("IOCs Deployed:", "=Program_KPIs!K13"),
        ("Active Exploits Addressed:", "=COUNTIF('[ISMS-IMP-A.5.7.2.xlsx]Vulnerability_Linked_Threats'!R:R,\"Patched\")"),
    ]
    
    for label, formula in metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"C{row}"] = formula
        ws[f"C{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1
    
    row += 1
    
    # Section 3: Integration Highlights
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "3. INTEGRATION & DISTRIBUTION"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    
    row += 1
    ws.merge_cells(f"A{row}:F{row+2}")
    integration_text = """Tools Integrated: [From Integration_Status]
IOC Hit Rate: [From Integration_Status]
Distribution Reach: [From Stakeholder_Engagement]"""
    ws[f"A{row}"] = integration_text
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws[f"A{row}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws.row_dimensions[row].height = 60
    
    row += 3
    
    # Section 4: Action Items
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "4. CRITICAL ACTION ITEMS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    
    row += 1
    ws.merge_cells(f"A{row}:F{row+3}")
    actions_text = """[User completes based on Risk_Indicators and open action items]

Critical Risks: [From Risk_Indicators Critical count]
Open High Priority Actions: [From source workbooks]
Recommendations: [User input]"""
    ws[f"A{row}"] = actions_text
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws[f"A{row}"].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws.row_dimensions[row].height = 80
    
    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 20


def create_quarterly_report(ws, styles):
    """Sheet 11: Quarterly Report - Comprehensive quarterly review."""
    
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "Threat Intelligence Program - Quarterly Report"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Report metadata
    ws.merge_cells("A2:C2")
    ws["A2"] = f"Report Period: Q{((datetime.now().month-1)//3)+1} {datetime.now().year}"
    ws["A2"].font = Font(name="Calibri", size=11, bold=True)
    
    ws.merge_cells("D2:F2")
    ws["D2"] = f"Generated: {datetime.now().strftime('%d.%m.%Y')}"
    ws["D2"].font = Font(name="Calibri", size=11, italic=True)
    ws["D2"].alignment = Alignment(horizontal="right")
    
    row = 4
    
    # Quarterly sections
    sections = [
        ("1. QUARTERLY EXECUTIVE SUMMARY", "70AD47"),
        ("2. STRATEGIC THREAT LANDSCAPE", "5B9BD5"),
        ("3. PROGRAM MATURITY ASSESSMENT", "4472C4"),
        ("4. QUARTERLY KPI PERFORMANCE", "FFC000"),
        ("5. SOURCE PORTFOLIO ROI ANALYSIS", "70AD47"),
        ("6. CONTROL 8.8 INTEGRATION REVIEW", "C00000"),
        ("7. STAKEHOLDER ENGAGEMENT TRENDS", "5B9BD5"),
        ("8. COMPLIANCE STATUS", "4472C4"),
        ("9. BUDGET VS. ACTUAL", "FFC000"),
        ("10. NEXT QUARTER PRIORITIES", "C00000"),
    ]
    
    for section, color in sections:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = section
        ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        row += 1
        ws.merge_cells(f"A{row}:F{row+2}")
        ws[f"A{row}"] = "[User completes with quarterly analysis and strategic insights]"
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"A{row}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws.row_dimensions[row].height = 60
        
        row += 3
    
    # Data aggregation note
    ws.merge_cells(f"A{row}:F{row+1}")
    note = """Note: This report requires manual completion with strategic insights.
Aggregate data from Trend_Analysis sheet for quarterly trends.
Reference Process_Maturity assessment from 5.7.2 Sheet 9."""
    ws[f"A{row}"] = note
    ws[f"A{row}"].font = Font(name="Calibri", size=9, italic=True)
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 40
    
    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 20


def create_metadata(ws, styles):
    """Sheet 12: Metadata - Dashboard generation information."""
    
    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Dashboard Metadata"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    generation_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    metadata = [
        ("Document ID:", "ISMS-IMP-A.5.7.4"),
        ("Document Title:", "Threat Intelligence Effectiveness Dashboard"),
        ("ISO 27001:2022 Control:", "A.5.7 (Threat Intelligence)"),
        ("Version:", "1.0"),
        ("Generated Date:", generation_date),
        ("Generator Script:", "generate_a57_4_dashboard.py"),
        ("Python Version:", "3.x"),
        ("Required Library:", "openpyxl"),
        ("", ""),
        ("Sheet Count:", "12"),
        ("Sheet 1:", "Executive_Summary (C-level dashboard)"),
        ("Sheet 2:", "Program_KPIs (20 standard KPIs with external refs)"),
        ("Sheet 3:", "Source_Portfolio (aggregated from 5.7.1)"),
        ("Sheet 4:", "Intelligence_Operations (aggregated from 5.7.2)"),
        ("Sheet 5:", "Integration_Status (aggregated from 5.7.3)"),
        ("Sheet 6:", "Stakeholder_Engagement (aggregated from 5.7.3)"),
        ("Sheet 7:", "Trend_Analysis (12-month historical tracking)"),
        ("Sheet 8:", "Risk_Indicators (early warning system)"),
        ("Sheet 9:", "Compliance_Evidence (20 audit artifacts)"),
        ("Sheet 10:", "Monthly_Report (executive summary)"),
        ("Sheet 11:", "Quarterly_Report (strategic review)"),
        ("Sheet 12:", "Metadata (this sheet)"),
        ("", ""),
        ("External References:", "CRITICAL - Must update file paths after generation"),
        ("Source Workbooks:", "5.7.1 (Sources), 5.7.2 (Collection), 5.7.3 (Integration)"),
        ("Integration Point:", "VTL metrics from 5.7.2 Sheet 8 (Control 8.8 linkage)"),
        ("", ""),
        ("Setup Instructions:", "1. Place all 5.7.x workbooks in same directory"),
        ("", "2. Open this dashboard"),
        ("", "3. Data → Edit Links → Update paths"),
        ("", "4. Click 'Update' when prompted to refresh external links"),
        ("", "5. Set target values in Program_KPIs sheet"),
        ("", "6. Configure risk thresholds in Risk_Indicators"),
        ("", ""),
        ("Review Cycle:", "Monthly refresh, Quarterly deep review"),
        ("Policy Reference:", "ISMS-POL-A.5.7 (All Sections)"),
        ("Related Workbooks:", "ISMS-IMP-A.5.7.1, ISMS-IMP-A.5.7.2, ISMS-IMP-A.5.7.3"),
        ("", ""),
        ("Notes:", "External references enable auto-aggregation from source workbooks"),
        ("", "Charts and visualizations require manual creation or BI tools"),
        ("", "As Feynman said: 'The first principle is that you must not fool yourself'"),
        ("", "This dashboard synthesizes evidence, not theater."),
    ]
    
    row = 3
    for label, value in metadata:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        
        if label in ["Document ID:", "Document Title:", "External References:", "Setup Instructions:"]:
            ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        elif "CRITICAL" in str(value):
            ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True, color="FF0000")
        else:
            ws[f"A{row}"].font = Font(name="Calibri", size=10)
            ws[f"B{row}"].font = Font(name="Calibri", size=10)
        
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Generate ISMS-IMP-A.5.7.4 dashboard workbook."""
    
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.5.7.4 V1.0 - Threat Intelligence Effectiveness Dashboard Generator (CVSS Enhanced)")
    logger.info("ISO/IEC 27001:2022 Control A.5.7 (Threat Intelligence)")
    logger.info("=" * 80)
    logger.info("")
    logger.info("⚠️  PREREQUISITE: Generate workbooks 5.7.1, 5.7.2, 5.7.3 FIRST")
    logger.info("⚠️  CRITICAL: External references must be updated after generation")
    logger.info("")
    
    # Create workbook
    logger.info("[1/13] Creating workbook structure...")
    wb = create_workbook()
    
    # Setup styles and validations
    logger.info("[2/13] Setting up styles and validations...")
    styles = setup_styles()
    validations = setup_validations()
    
    # Create sheets
    logger.info("[3/13] Creating Sheet 1: Executive_Summary (Dashboard)...")
    create_executive_summary(wb["Executive_Summary"], styles)
    
    logger.info("[4/13] Creating Sheet 2: Program_KPIs...")
    create_program_kpis(wb["Program_KPIs"], styles, validations)
    
    logger.info("[5/13] Creating Sheet 3: Source_Portfolio...")
    create_source_portfolio(wb["Source_Portfolio"], styles)
    
    logger.info("[6/13] Creating Sheet 4: Intelligence_Operations...")
    create_intelligence_operations(wb["Intelligence_Operations"], styles)
    
    logger.info("[7/13] Creating Sheet 5: Integration_Status...")
    create_integration_status(wb["Integration_Status"], styles)
    
    logger.info("[8/13] Creating Sheet 6: Stakeholder_Engagement...")
    create_stakeholder_engagement(wb["Stakeholder_Engagement"], styles)
    
    logger.info("[9/13] Creating Sheet 7: Trend_Analysis...")
    create_trend_analysis(wb["Trend_Analysis"], styles)
    
    logger.info("[10/13] Creating Sheet 8: Risk_Indicators...")
    create_risk_indicators(wb["Risk_Indicators"], styles, validations)
    
    logger.info("[11/13] Creating Sheet 9: Compliance_Evidence...")
    create_compliance_evidence(wb["Compliance_Evidence"], styles, validations)
    
    logger.info("[12/13] Creating Sheet 10: Monthly_Report...")
    create_monthly_report(wb["Monthly_Report"], styles)
    
    logger.info("[13/13] Creating Sheet 11: Quarterly_Report...")
    create_quarterly_report(wb["Quarterly_Report"], styles)
    
    logger.info("[14/13] Creating Sheet 12: Metadata...")
    create_metadata(wb["Metadata"], styles)
    
    # Generate filename with date
    filename = f"ISMS-IMP-A.5.7.4_Dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    logger.info("")
    logger.info(f"[15/13] Saving workbook: {filename}")
    wb.save(filename)
    
    logger.info("")
    logger.info("=" * 80)
    logger.info("✅ SUCCESS!")
    logger.info("=" * 80)
    logger.info("")
    logger.info(f"Generated: {filename}")
    logger.info("")
    logger.info("Workbook Contents:")
    logger.info("  - 12 sheets with comprehensive TI program dashboard")
    logger.info("  - Executive summary (single-page C-level view)")
    logger.info("  - 24 KPIs: 18 standard + 6 CVSS/Risk/Incident KPIs (V1.0)")
    logger.info("  - Aggregated metrics from 5.7.1, 5.7.2, 5.7.3 workbooks")
    logger.info("  - 12-month trend analysis")
    logger.info("  - Risk indicators and early warnings")
    logger.info("  - 20 compliance evidence items for ISO 27001 audits")
    logger.info("  - Monthly and quarterly report templates")
    logger.info("")
    logger.info("⚠️  CRITICAL POST-GENERATION STEPS:")
    logger.info("=" * 80)
    logger.info("1. Complete assessment workbooks with operational data:")
    logger.info("   - ISMS-IMP-A.5.7.1_Sources_YYYYMMDD.xlsx")
    logger.info("   - ISMS-IMP-A.5.7.2_Collection_YYYYMMDD.xlsx")
    logger.info("   - ISMS-IMP-A.5.7.3_Integration_YYYYMMDD.xlsx")
    logger.info("")
    logger.info("2. Normalize filenames for stable dashboard references:")
    logger.info("   python3 normalize_assessment_files_a57.py")
    logger.info("   → Creates: ISMS-IMP-A.5.7.1.xlsx, ISMS-IMP-A.5.7.2.xlsx, ISMS-IMP-A.5.7.3.xlsx")
    logger.info("   → Location: ./Dashboard_Sources/ directory")
    logger.info("")
    logger.info("3. Place this dashboard in Dashboard_Sources/ directory:")
    logger.info("   cp ISMS-IMP-A.5.7.4_Dashboard_YYYYMMDD.xlsx Dashboard_Sources/")
    logger.info("")
    logger.info("4. Open dashboard in Excel:")
    logger.info("   - Excel prompts: 'Update Links?' → Click UPDATE")
    logger.info("   - Dashboard auto-populates with current TI program data")
    logger.info("")
    logger.info("5. Complete dashboard setup:")
    logger.info("   - Set target values in Program_KPIs sheet")
    logger.info("   - Configure risk thresholds in Risk_Indicators")
    logger.info("   - Update compliance evidence status")
    logger.info("")
    logger.info("External References Used:")
    logger.info("  - [ISMS-IMP-A.5.7.1.xlsx] → 50+ formulas (normalized, no dates)")
    logger.info("  - [ISMS-IMP-A.5.7.2.xlsx] → 40+ formulas (normalized, no dates)")
    logger.info("  - [ISMS-IMP-A.5.7.3.xlsx] → 30+ formulas (normalized, no dates)")
    logger.info("")
    logger.info("Next Steps:")
    logger.info("  1. Run sanity check: python3 excel_sanity_check_a57_4.py <filename>")
    logger.info("  2. Normalize assessment files: python3 normalize_assessment_files_a57.py")
    logger.info("  3. Place dashboard in Dashboard_Sources/ directory")
    logger.info("  4. Open dashboard and click 'Update Links'")
    logger.info("")
    
    
    logger.info("")
    logger.info("NEW IN V1.0:")
    logger.info("  ✓ CVSS 4.0 adoption rate tracking (target: 75% by Q4 2026)")
    logger.info("  ✓ 6 CVSS-based KPIs (CVSS-001 to INC-001)")
    logger.info("  ✓ Emergency indicator (High CVSS + Active Exploitation = 0)")
    logger.info("  ✓ CVSS severity distribution of prevented incidents")
    logger.info("  ✓ CVSS quantification for risk management (Clause 6.1)")
    logger.info("  ✓ CVSS context in incident response tracking")
    logger.info("  ✓ Source CVSS accuracy monitoring")
    logger.info("")
    logger.info("CRITICAL: Ensure source workbooks (A.5.7.1/2/3) have CVSS columns!")
    
    logger.info("Evidence > Theater. Use it well.")
    logger.info("=" * 80)


if __name__ == "__main__":
    main()


# ============================================================================
# END OF GENERATE_A57_4_DASHBOARD.PY
# ============================================================================
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
