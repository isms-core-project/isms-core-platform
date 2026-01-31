#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.7.2 - Intelligence Collection & Analysis Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.7: Threat Intelligence
Assessment Domain 2 of 5: Intelligence Collection & Analysis Workflows

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific threat intelligence workflows, analysis processes,
and operational requirements.

Key customization areas:
1. Intelligence collection workflows and automation (match your SIEM/SOAR tools)
2. Analysis team structure and responsibilities (adapt to your org chart)
3. MITRE ATT&CK coverage requirements (align with your threat model)
4. Indicator processing volumes and SLAs (based on operational capacity)
5. Quality assurance thresholds (confidence levels, validation requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.7 Threat Intelligence Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
threat intelligence collection, processing, analysis, and quality assurance
workflows against ISO 27001:2022 Control A.5.7 requirements.

**Purpose:**
Enables systematic assessment of how threat intelligence is collected from
sources (assessed in A.5.7.1), analyzed by security teams, enriched with
context, and prepared for operational use in security workflows.

**Assessment Scope:**
- Automated intelligence collection mechanisms (API integrations, feeds)
- Manual intelligence collection workflows (OSINT, vendor portals)
- Indicator processing and normalization (IOCs, TTPs, CVSS enrichment)
- Threat analysis workflows and analyst capabilities
- MITRE ATT&CK framework mapping and coverage
- Intelligence enrichment and contextualization
- Quality assurance and validation processes
- Analyst productivity and workload management
- Intelligence aging and archival policies
- Integration with SIEM, SOAR, and security tools
- CVSS validation and accuracy improvement (VTL support)
- Audit evidence collection and metrics

**Generated Workbook Structure (12 Sheets):**
1. Instructions - Assessment guidance and workflow methodology
2. Collection_Workflows - Automated and manual collection processes
3. Processing_Pipeline - Indicator normalization and enrichment
4. Analysis_Team - Analyst capabilities, training, and workload
5. MITRE_Coverage - ATT&CK technique coverage and gap analysis
6. Quality_Assurance - Validation processes and accuracy metrics
7. Tool_Integration - SIEM/SOAR/security tool integrations
8. Performance_Metrics - SLA compliance, processing volumes, timeliness
9. Intelligence_Lifecycle - Aging, archival, and retention policies
10. Gap_Analysis - Workflow deficiencies and remediation plans
11. Evidence_Register - Audit evidence tracking
12. Metadata - Assessment metadata and approval workflow

**Key Features:**
- MITRE ATT&CK technique coverage heatmap with gap visualization
- CVSS enrichment and validation workflow tracking (VTL integration)
- Analyst workload analysis and capacity planning
- Indicator processing SLA compliance monitoring
- Data validation with industry-standard dropdown lists
- Conditional formatting for workflow health status
- Protected formulas with unprotected input cells
- Integration with A.5.7.1 (source utilization) and A.5.7.3 (distribution)
- Evidence linkage for audit traceability

**VulnerabilityThreatLink (VTL) Integration:**
This assessment tracks CVSS enrichment and validation workflows that ensure
threat intelligence includes accurate severity ratings. Quality CVSS data
from collection/analysis enables effective vulnerability-threat correlation
and risk-based patching decisions in Control A.8.8.

**Integration:**
This assessment feeds into:
- A.5.7.4 Threat Intelligence Effectiveness Dashboard (workflow metrics)
- A.5.7.5 Standalone Compliance Dashboard (executive reporting)
- A.5.7.1 Sources Assessment (source utilization and effectiveness)
- A.5.7.3 Integration & Distribution (consumer requirements)
- A.8.8 Vulnerability Management (CVSS quality for patch prioritization)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a57_2_collection_analysis.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a57_2_collection_analysis.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a57_2_collection_analysis.py --date 20250115

Output:
    File: ISMS_A_5_7_2_Collection_Analysis_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)
    Sheets: 12 worksheets (see structure above)

Post-Generation Steps:
    1. Document all automated collection workflows (API integrations, feeds)
    2. Map manual collection processes and analyst responsibilities
    3. Assess indicator processing pipeline and normalization quality
    4. Evaluate analyst team capabilities, training, and workload
    5. Complete MITRE ATT&CK coverage analysis and gap identification
    6. Review quality assurance processes and validation metrics
    7. Verify tool integrations (SIEM, SOAR, ticketing, etc.)
    8. Measure performance against SLAs (timeliness, accuracy, volume)
    9. Define intelligence lifecycle policies (aging, archival, retention)
    10. Identify workflow gaps and create remediation plans
    11. Collect audit evidence (logs, metrics, validation reports)
    12. Obtain stakeholder approvals
    13. Feed results into A.5.7.4 Effectiveness Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.7
Assessment Domain:    2 of 5 (Intelligence Collection & Analysis Workflows)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.7: Threat Intelligence Policy (Governance)
    - ISMS-IMP-A.5.7.2: Intelligence Collection & Analysis Implementation Guide
    - ISMS-IMP-A.5.7.1: Threat Intelligence Sources Assessment (Domain 1)
    - ISMS-IMP-A.5.7.3: Intelligence Integration & Distribution Assessment (Domain 3)
    - ISMS-IMP-A.5.7.4: Threat Intelligence Effectiveness Dashboard (Consolidation)
    - ISMS-IMP-A.5.7.5: Standalone Compliance Dashboard (Executive Reporting)
    - ISO 27002:2022 Implementation Guidance for Control A.5.7
    - MITRE ATT&CK Framework (Enterprise v15.1)
    - STIX/TAXII Standards for Threat Intelligence Exchange
    - CVSS Scoring Standard v4.0 / v3.1 Specifications

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
- Expanded MITRE ATT&CK coverage analysis (Enterprise v15.1)
    - Added CVSS enrichment and validation workflow tracking (VTL integration)
    - Enhanced quality assurance metrics (confidence levels, false positives)
    - Added analyst workload and capacity planning analysis
    - Improved integration tracking with SIEM/SOAR platforms
    - Added intelligence lifecycle management assessment
    - Enhanced audit evidence collection mechanisms
    - Updated conditional formatting for workflow health visualization

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Intelligence Quality Standards:**
Collection and analysis workflows directly impact intelligence quality.
Establish clear quality metrics:
- Accuracy: Validation rate, false positive/negative percentages
- Timeliness: Time from source ingestion to analyst review to distribution
- Completeness: CVSS enrichment rate, context completeness, attribution quality
- Consistency: Normalization success rate, schema compliance
- Actionability: Percentage of intelligence leading to security actions

Poor workflow quality degrades downstream security operations effectiveness.

**MITRE ATT&CK Coverage (AUDIT CRITICAL):**
Auditors will verify that threat intelligence collection covers relevant
techniques based on your threat model. Document:
- Coverage percentage by ATT&CK tactic (Reconnaissance through Impact)
- Justification for gaps (techniques irrelevant to your environment)
- Plans to address coverage deficiencies
- Regular review cycle (quarterly recommended)

Incomplete coverage may indicate blind spots in threat detection capabilities.

**Analyst Workload Management:**
Unsustainable analyst workload leads to:
- Decreased analysis quality (rushed assessments)
- Increased analyst burnout and turnover
- Processing backlogs and SLA violations
- Missed critical intelligence

Monitor analyst productivity metrics and adjust staffing/automation accordingly.
Target: <40 hours/week analysis workload per analyst for sustainable operations.

**CVSS Enrichment Workflow (VTL Critical):**
For threat intelligence lacking CVSS scores, document enrichment processes:
- Who performs CVSS scoring (analyst roles, training)
- Validation against authoritative sources (NVD, vendor advisories)
- Quality assurance processes (peer review, spot checks)
- Accuracy targets (>90% agreement with authoritative scores)

CVSS enrichment quality directly impacts vulnerability management automation.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Auditors will verify:
- Documented collection workflows (automated and manual)
- Analyst competency and training records
- Quality assurance processes and validation results
- Integration with security operations workflows
- Performance metrics and SLA compliance

Maintain evidence in Sheet 11 (Evidence_Register) with proper linkage.

**Data Protection:**
Assessment workbooks contain sensitive operational details:
- Threat intelligence workflows and analysis procedures
- Analyst identities and capabilities assessment
- Tool integration technical details
- Performance metrics revealing operational gaps

Classification: CONFIDENTIAL - Internal Use Only. Restrict to security team.

**Maintenance:**
Review and update assessment:
- Monthly: Performance metrics and SLA compliance
- Quarterly: MITRE ATT&CK coverage review and analyst workload analysis
- Semi-annually: Workflow optimization and tool integration review
- Annually: Complete reassessment of collection and analysis capabilities
- Ad-hoc: When workflows change or new tools are deployed

**Quality Assurance:**
Have senior threat intelligence analysts and SOC leadership validate
assessments before using results for resource allocation or workflow changes.
Cross-reference with A.5.7.1 (source effectiveness) and A.5.7.3 (consumer needs).

**Regulatory Alignment:**
Some regulations require specific threat intelligence capabilities:
- Financial sector: Fraud and financial crime intelligence workflows
- Healthcare: Healthcare-specific threat intelligence (medical device, EHR)
- Critical infrastructure: ICS/SCADA threat intelligence processes
- Government: Classified intelligence handling procedures (if applicable)

Customize workflow assessments to include regulatory-specific requirements.

**Tool Integration Dependencies:**
Collection and analysis effectiveness depends heavily on tool integrations:
- SIEM: Automated indicator ingestion and correlation
- SOAR: Playbook automation for intelligence-driven response
- Threat intelligence platform (TIP): Central intelligence repository
- Vulnerability scanners: CVSS correlation for patch prioritization
- EDR/XDR: Endpoint indicator deployment and hunting

Poor integrations create manual workflows that don't scale. Prioritize automation.

**MITRE ATT&CK Framework Usage:**
This assessment uses MITRE ATT&CK Enterprise Framework v15.1 (or current).
Update technique coverage mapping when new ATT&CK versions are released.
Focus coverage analysis on tactics/techniques relevant to your threat model.

**Business Impact:**
Inadequate collection and analysis workflows result in:
- Delayed threat detection and incident response
- Missed threat actor campaigns targeting your sector
- Inability to prioritize vulnerabilities based on exploitation likelihood
- Reduced security ROI from threat intelligence investments

This assessment helps quantify and remediate workflow deficiencies.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)




# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.7.2"
WORKBOOK_NAME = "Intelligence Collection & Analysis Assessment"
CONTROL_ID = "A.5.7"
CONTROL_NAME = "Threat Intelligence"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching specification."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure v1.0: 14 sheets (aligned with IMP specification)
    # Updated to match ISMS-IMP-A.5.7.2 for industry best practices:
    # - Intelligence_Requirements (PIR/SIR framework)
    # - Collection_Sources (data source documentation)
    # - Raw_Intelligence_Log (tracking raw intel)
    # - Process_Maturity (capability maturity assessment)
    sheets = [
        "Instructions",
        "Intelligence_Requirements",     # PIR/SIR/IR tracking
        "Collection_Sources",            # Source documentation
        "Raw_Intelligence_Log",          # Raw intel tracking
        "Intelligence_Production",
        "MITRE_Mapping",
        "Quality_Metrics",
        "Vulnerability_Linked_Threats",  # CRITICAL: VTL with CVSS (v1.0)
        "Process_Maturity",              # Capability maturity assessment
        "Action_Items",
        "Analysis_Tools",
        "Threat_Actor_Profiles",
        "Campaign_Tracking",
        "Metadata",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="8FAADC", end_color="8FAADC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "formula_cell": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
    }
    return styles


def get_border():
    """Create new border object (avoid shared object warnings)."""
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATION FUNCTIONS
# ============================================================================

def setup_validations():
    """Create all data validation dropdowns."""
    validations = {}

    # Intelligence Type
    validations['intelligence_type'] = DataValidation(
        type="list",
        formula1='"Strategic,Tactical,Operational,Technical"',
        allow_blank=False
    )
    
    # Priority
    validations['priority'] = DataValidation(
        type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚫ Low"',
        allow_blank=False
    )
    
    # Collection Method
    validations['collection_method'] = DataValidation(
        type="list",
        formula1='"Automated_Feed,Manual_Research,OSINT,Partnership,Internal_Telemetry"',
        allow_blank=False
    )
    
    # Collection Frequency
    validations['collection_frequency'] = DataValidation(
        type="list",
        formula1='"Real-time,Hourly,Daily,Weekly,On_Demand"',
        allow_blank=False
    )
    
    # Framework Type
    validations['framework_type'] = DataValidation(
        type="list",
        formula1='"Threat_Model,Analysis_Method,Taxonomy,Tool"',
        allow_blank=False
    )
    
    # Use Case
    validations['use_case'] = DataValidation(
        type="list",
        formula1='"Threat_Profiling,IOC_Analysis,Campaign_Tracking,Reporting,Integration"',
        allow_blank=False
    )
    
    # Implementation Status
    validations['implementation_status'] = DataValidation(
        type="list",
        formula1='"Fully_Implemented,Partially_Implemented,Planned,Not_Used"',
        allow_blank=False
    )
    
    # Yes/No/Optional
    validations['yes_no_optional'] = DataValidation(
        type="list",
        formula1='"Yes,No,Optional"',
        allow_blank=False
    )
    
    # Effectiveness Rating
    validations['effectiveness_rating'] = DataValidation(
        type="list",
        formula1='"Excellent,Good,Fair,Poor"',
        allow_blank=True
    )
    
    # Role
    validations['analyst_role'] = DataValidation(
        type="list",
        formula1='"Lead_Analyst,Senior_Analyst,Analyst,Junior_Analyst,Intern"',
        allow_blank=False
    )
    
    # Primary Focus Area
    validations['focus_area'] = DataValidation(
        type="list",
        formula1='"Strategic,Tactical,Operational,Technical"',
        allow_blank=False
    )
    
    # Skill Levels
    validations['skill_level'] = DataValidation(
        type="list",
        formula1='"Expert,Advanced,Intermediate,Beginner,None"',
        allow_blank=False
    )
    
    # Product Type
    validations['product_type'] = DataValidation(
        type="list",
        formula1='"Strategic_Report,Tactical_Brief,Operational_Alert,Technical_Analysis,IOC_Package,Threat_Profile"',
        allow_blank=False
    )
    
    # Product Status
    validations['product_status'] = DataValidation(
        type="list",
        formula1='"Draft,Review,Published,Archived"',
        allow_blank=False
    )
    
    # Confidence Level
    validations['confidence_level'] = DataValidation(
        type="list",
        formula1='"High,Medium,Low,Unconfirmed"',
        allow_blank=False
    )
    
    # TLP Classification
    validations['tlp'] = DataValidation(
        type="list",
        formula1='"TLP:CLEAR,TLP:GREEN,TLP:AMBER,TLP:AMBER+STRICT,TLP:RED"',
        allow_blank=False
    )
    
    # MITRE Coverage
    validations['mitre_coverage'] = DataValidation(
        type="list",
        formula1='"Full,Partial,Limited,None"',
        allow_blank=False
    )
    
    # Threat Actor Type (VTL schema)
    validations['threat_actor_type'] = DataValidation(
        type="list",
        formula1='"Nation-State,Criminal_Syndicate,Hacktivist,Insider_Threat,Script_Kiddie,Unknown"',
        allow_blank=True
    )
    
    # Exploitation Status (VTL schema - CRITICAL)
    validations['exploitation_status'] = DataValidation(
        type="list",
        formula1='"Not_Exploited,PoC_Exists,Active_Exploitation,Widespread"',
        allow_blank=False
    )
    
    # Remediation Urgency (VTL schema)
    validations['remediation_urgency'] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low,Info"',
        allow_blank=False
    )
    
    # Remediation Status (VTL schema)
    validations['remediation_status'] = DataValidation(
        type="list",
        formula1='"Open,In_Progress,Patched,Mitigated,Risk_Accepted,Verified"',
        allow_blank=False
    )
    
    # Yes/No
    validations['yes_no'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    
    # Maturity Levels (1-5 CMM)
    validations['maturity_level'] = DataValidation(
        type="list",
        formula1='"1 - Initial,2 - Managed,3 - Defined,4 - Quantitatively Managed,5 - Optimizing"',
        allow_blank=False
    )
    
    # Action Status
    validations['action_status'] = DataValidation(
        type="list",
        formula1='"Open,In_Progress,Blocked,Resolved,Closed"',
        allow_blank=False
    )
    

    # NEW v1.0 validations for Sheets 11-13
    validations['tool_category'] = DataValidation(
        type="list",
        formula1='"TIP,SIEM,Malware_Analysis,OSINT,Collaboration,Visualization,Scripting,Other"',
        allow_blank=True
    )
    validations['license_type'] = DataValidation(
        type="list",
        formula1='"Commercial,Open_Source,Freeware,Internal_Developed"',
        allow_blank=True
    )
    validations['yes_no_na'] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=True
    )
    validations['yes_no_unknown'] = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=True
    )
    validations['targeting_status'] = DataValidation(
        type="list",
        formula1='"Confirmed,Suspected,No,Unknown"',
        allow_blank=True
    )
    validations['motivation'] = DataValidation(
        type="list",
        formula1='"Espionage,Financial,Disruption,Ideology,Unknown"',
        allow_blank=True
    )
    validations['sophistication'] = DataValidation(
        type="list",
        formula1='"Advanced,Moderate,Low"',
        allow_blank=True
    )
    validations['campaign_status'] = DataValidation(
        type="list",
        formula1='"Active,Concluded,Dormant,Unknown"',
        allow_blank=True
    )
    validations['campaign_objective'] = DataValidation(
        type="list",
        formula1='"Espionage,Data_Theft,Disruption,Ransomware,Credential_Harvesting,Other"',
        allow_blank=True
    )
    validations['threat_level'] = DataValidation(
        type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚫ Low"',
        allow_blank=True
    )
    validations['monitoring_status'] = DataValidation(
        type="list",
        formula1='"Active_Monitoring,Passive_Monitoring,Concluded"',
        allow_blank=True
    )

    validations['cvss_version'] = DataValidation(
        type="list",
        formula1='"4.0,3.1"',
        allow_blank=True
    )
    

    # Validations for Sheet 11 (Analysis_Tools)
    validations['tool_category'] = DataValidation(
        type="list",
        formula1='"TIP,SIEM,Malware_Analysis,OSINT,Collaboration,Visualization,Scripting,Other"',
        allow_blank=True
    )
    validations['license_type'] = DataValidation(
        type="list",
        formula1='"Commercial,Open_Source,Freeware,Internal_Developed"',
        allow_blank=True
    )
    validations['integration_status'] = DataValidation(
        type="list",
        formula1='"Integrated,Standalone,Planned,Deprecated"',
        allow_blank=True
    )
    validations['yes_no_na'] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=True
    )
    
    # Validations for Sheet 12 (Threat_Actor_Profiles)
    validations['yes_no_unknown'] = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=True
    )
    validations['targeting_status'] = DataValidation(
        type="list",
        formula1='"Confirmed,Suspected,No,Unknown"',
        allow_blank=True
    )
    validations['motivation'] = DataValidation(
        type="list",
        formula1='"Espionage,Financial,Disruption,Ideology,Unknown"',
        allow_blank=True
    )
    validations['sophistication'] = DataValidation(
        type="list",
        formula1='"Advanced,Moderate,Low"',
        allow_blank=True
    )
    
    # Validations for Sheet 13 (Campaign_Tracking)
    validations['campaign_status'] = DataValidation(
        type="list",
        formula1='"Active,Concluded,Dormant,Unknown"',
        allow_blank=True
    )
    validations['campaign_objective'] = DataValidation(
        type="list",
        formula1='"Espionage,Data_Theft,Disruption,Ransomware,Credential_Harvesting,Other"',
        allow_blank=True
    )
    validations['threat_level'] = DataValidation(
        type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚫ Low"',
        allow_blank=True
    )
    validations['monitoring_status'] = DataValidation(
        type="list",
        formula1='"Active_Monitoring,Passive_Monitoring,Concluded"',
        allow_blank=True
    )
    

    return validations


# ============================================================================
# SECTION 3: SHEET 1 - INSTRUCTIONS
# ============================================================================

def create_instructions(ws, styles):
    """Create Instructions sheet with completion guidance."""
    
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "ISMS-IMP-A.5.7.2 - Intelligence Collection & Analysis Assessment"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    # Subtitle
    ws.merge_cells("A2:G2")
    ws["A2"] = "Assessment of Intelligence Collection, Analysis Capabilities, and Control 8.8 Integration"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    
    row = 4
    
    # Purpose section
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "PURPOSE & OBJECTIVES"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    
    row += 1
    ws.merge_cells(f"A{row}:G{row+3}")
    purpose_text = """This workbook assesses the organization's capability to collect, analyze, and produce actionable threat intelligence:
- Intelligence collection coverage against requirements
- Analysis frameworks and methodologies (MITRE ATT&CK, Diamond Model, etc.)
- Analyst skills and capability development
- Intelligence production quality and timeliness
- **CRITICAL: VulnerabilityThreatLink (VTL) schema for Control 8.8 integration**"""
    ws[f"A{row}"] = purpose_text
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 80
    
    row += 4
    
    # Sheet Guide section
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "SHEET GUIDE"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    
    row += 1
    sheet_guide = [
        ("Sheet 2", "Intelligence_Requirements", "Document PIRs, SIRs, and IRs per best practices"),
        ("Sheet 3", "Collection_Sources", "Document intelligence collection sources and methods"),
        ("Sheet 4", "Raw_Intelligence_Log", "Track raw intelligence received and processed"),
        ("Sheet 5", "Intelligence_Production", "Track intelligence products and quality"),
        ("Sheet 6", "MITRE_Mapping", "ATT&CK technique coverage analysis"),
        ("Sheet 7", "Quality_Metrics", "KPIs and performance tracking"),
        ("Sheet 8", "Vulnerability_Linked_Threats", "**VTL SCHEMA - Control 8.8 Integration**"),
        ("Sheet 9", "Process_Maturity", "CMM capability maturity assessment"),
        ("Sheet 10", "Action_Items", "Capability improvement tracking"),
        ("Sheet 11", "Analysis_Tools", "Document analytical tools and platforms"),
        ("Sheet 12", "Threat_Actor_Profiles", "Track threat actors and campaigns"),
        ("Sheet 13", "Campaign_Tracking", "Document active campaigns and TTPs"),
        ("Sheet 14", "Metadata", "Workbook generation information"),
    ]
    
    for sheet_num, sheet_name, description in sheet_guide:
        ws[f"A{row}"] = sheet_num
        ws[f"B{row}"] = sheet_name
        ws[f"C{row}"] = description
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"].font = Font(name="Calibri", size=10)
        ws[f"C{row}"].font = Font(name="Calibri", size=10)
        if "VTL SCHEMA" in description:
            ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True, color="FF0000")
            ws[f"C{row}"].font = Font(name="Calibri", size=10, bold=True, color="FF0000")
        row += 1
    
    row += 1
    
    # CRITICAL Integration Notice
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "⚠️ CRITICAL: CONTROL 8.8 INTEGRATION"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FF0000")
    ws[f"A{row}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    ws.merge_cells(f"A{row}:G{row+4}")
    integration_notice = """Sheet 8 (Vulnerability_Linked_Threats) implements the VulnerabilityThreatLink (VTL) schema for bidirectional integration between:
- Control 5.7 (Threat Intelligence) → Threat context for vulnerability prioritization
- Control 8.8 (Vulnerability Management) → Vulnerability data for threat analysis

When threat intelligence identifies active exploitation of vulnerabilities:
1. Create VTL record in Sheet 8
2. VTL record syncs to Control 8.8 workbook (emergency patching trigger)
3. Remediation status updates flow back from 8.8 to 5.7"""
    ws[f"A{row}"] = integration_notice
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 100
    
    row += 5
    
    # Completion Instructions
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "COMPLETION INSTRUCTIONS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    
    row += 1
    ws.merge_cells(f"A{row}:G{row+7}")
    completion_text = """1. Intelligence_Requirements: Define PIRs, SIRs, IRs per industry best practices
2. Collection_Sources: Document internal/external collection sources and methods
3. Raw_Intelligence_Log: Track raw intelligence received and processing status
4. Intelligence_Production: Track intelligence products from last 90 days
5. MITRE_Mapping: Assess coverage of critical ATT&CK techniques
6. Quality_Metrics: Establish baseline KPIs and targets
7. Vulnerability_Linked_Threats: Create VTL records for exploited vulnerabilities
8. Process_Maturity: Assess capability maturity using CMM model
9. Action_Items: Document improvement initiatives"""
    ws[f"A{row}"] = completion_text
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 140
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15


# ============================================================================
# END OF SECTION 1
# ============================================================================

# ============================================================================
# SECTION 2: SHEETS 2-5 (COLLECTION & ANALYSIS)
# ============================================================================

def create_intelligence_requirements(ws, styles, validations):
    """Sheet 2: Intelligence_Requirements - PIR/SIR/IR tracking per IMP spec."""

    # Title
    ws.merge_cells("A1:R1")
    ws["A1"] = "Intelligence Requirements (PIR/SIR/IR)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:R2")
    ws["A2"] = "Document all intelligence requirements (PIRs, SIRs, IRs). Map sources from A.5.7.1. Identify coverage gaps."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Column headers (Row 4)
    headers = [
        "Requirement_ID", "Intelligence_Type", "Requirement_Description", "Priority",
        "Target_Sector", "Target_Region", "Threat_Category", "Collection_Source_1",
        "Collection_Source_2", "Collection_Source_3", "Collection_Method",
        "Coverage_Status", "Collection_Frequency", "Last_Collected", "Gap_Identified",
        "Gap_Remediation", "Responsible_Analyst", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 30
    
    # Data rows (50 requirements capacity: rows 5-54)
    for row in range(5, 55):
        # Requirement_ID (auto-generated)
        cell = ws.cell(row=row, column=1)
        cell.value = f"REQ-2025-{str(row-4).zfill(3)}"
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        
        # Intelligence_Type
        cell = ws.cell(row=row, column=2)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['intelligence_type'].add(cell)
        
        # Requirement_Description
        cell = ws.cell(row=row, column=3)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # Priority
        cell = ws.cell(row=row, column=4)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['priority'].add(cell)
        
        # Target_Sector, Target_Region, Threat_Category (text input)
        for col in [5, 6, 7]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Collection_Source_1, 2, 3 (text input - reference to 5.7.1)
        for col in [8, 9, 10]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Collection_Method
        cell = ws.cell(row=row, column=11)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['collection_method'].add(cell)
        
        # Coverage_Status (formula)
        cell = ws.cell(row=row, column=12)
        cell.value = f'=IF(COUNTA(H{row}:J{row})>=2,"Adequate",IF(COUNTA(H{row}:J{row})=1,"Minimal","Gap"))'
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        
        # Collection_Frequency
        cell = ws.cell(row=row, column=13)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['collection_frequency'].add(cell)
        
        # Last_Collected (date)
        cell = ws.cell(row=row, column=14)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'
        
        # Gap_Identified (formula)
        cell = ws.cell(row=row, column=15)
        cell.value = f'=IF(L{row}="Gap","Yes","No")'
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        
        # Gap_Remediation, Responsible_Analyst, Notes
        for col in [16, 17, 18]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
    
    # Apply validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        'L5:L54',
        CellIsRule(operator='equal', formula=['"Gap"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'L5:L54',
        CellIsRule(operator='equal', formula=['"Minimal"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'L5:L54',
        CellIsRule(operator='equal', formula=['"Adequate"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    
    # Summary dashboard (row 56)
    ws.merge_cells("A56:D56")
    ws["A56"] = "COLLECTION COVERAGE SUMMARY"
    ws["A56"].font = Font(name="Calibri", size=11, bold=True)
    ws["A56"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A56"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    summary_metrics = [
        ("Total Requirements:", '=COUNTA(A5:A54)'),
        ("Critical Priority:", '=COUNTIF(D5:D54,"Critical")'),
        ("High Priority:", '=COUNTIF(D5:D54,"High")'),
        ("Coverage Gaps:", '=COUNTIF(L5:L54,"Gap")'),
        ("Minimal Coverage:", '=COUNTIF(L5:L54,"Minimal")'),
        ("Adequate Coverage:", '=COUNTIF(L5:L54,"Adequate")'),
    ]
    
    row = 57
    for label, formula in summary_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    for col in ['H', 'I', 'J']:
        ws.column_dimensions[col].width = 15
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 18
    ws.column_dimensions['N'].width = 14
    ws.column_dimensions['O'].width = 14
    ws.column_dimensions['P'].width = 30
    ws.column_dimensions['Q'].width = 25
    ws.column_dimensions['R'].width = 30
    
    ws.freeze_panes = "A5"


def create_collection_sources(ws, styles, validations):
    """Sheet 3: Collection_Sources - Document intelligence collection sources per IMP spec."""

    # Title
    ws.merge_cells("A1:P1")
    ws["A1"] = "Collection Sources"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    # Instructions
    ws.merge_cells("A2:P2")
    ws["A2"] = "Document collection workflows for each source from A.5.7.1. Map data formats, update frequency, and integration status."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Column headers per IMP specification
    headers = [
        "Source_ID", "Source_Name", "Source_Type", "Data_Format", "Update_Frequency",
        "Coverage_Geographic", "Coverage_Sector", "Coverage_Threat_Types",
        "Integration_Status", "Integration_Platform", "Cost_Annual", "Contract_Expiry",
        "Primary_Contact", "Data_Quality_Rating", "Last_Review_Date", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 30
    
    # Data rows (50 source capacity: rows 5-54)
    for row in range(5, 55):
        # Source_ID (col 1 - auto-generated)
        cell = ws.cell(row=row, column=1)
        cell.value = f"SRC-{str(row-4).zfill(3)}"
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()

        # Source_Name (col 2)
        cell = ws.cell(row=row, column=2)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Source_Type (col 3)
        cell = ws.cell(row=row, column=3)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Data_Format (col 4)
        cell = ws.cell(row=row, column=4)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Update_Frequency (col 5)
        cell = ws.cell(row=row, column=5)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['collection_frequency'].add(cell)

        # Coverage_Geographic (col 6)
        cell = ws.cell(row=row, column=6)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Coverage_Sector (col 7)
        cell = ws.cell(row=row, column=7)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Coverage_Threat_Types (col 8)
        cell = ws.cell(row=row, column=8)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Integration_Status (col 9)
        cell = ws.cell(row=row, column=9)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['implementation_status'].add(cell)

        # Integration_Platform (col 10)
        cell = ws.cell(row=row, column=10)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Cost_Annual (col 11)
        cell = ws.cell(row=row, column=11)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '#,##0.00'

        # Contract_Expiry (col 12)
        cell = ws.cell(row=row, column=12)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'

        # Primary_Contact (col 13)
        cell = ws.cell(row=row, column=13)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Data_Quality_Rating (col 14)
        cell = ws.cell(row=row, column=14)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['effectiveness_rating'].add(cell)

        # Last_Review_Date (col 15)
        cell = ws.cell(row=row, column=15)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'

        # Notes (col 16)
        cell = ws.cell(row=row, column=16)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

    # Apply validations
    for dv in validations.values():
        ws.add_data_validation(dv)

    # Summary dashboard (row 56)
    ws.merge_cells("A56:D56")
    ws["A56"] = "COLLECTION SOURCES SUMMARY"
    ws["A56"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A56"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    summary = [
        ("Total Sources:", '=COUNTA(A5:A54)'),
        ("Integrated Sources:", '=COUNTIF(I5:I54,"Fully_Implemented")'),
        ("Manual Sources:", '=COUNTIF(I5:I54,"Partially_Implemented")'),
        ("Excellent Quality:", '=COUNTIF(N5:N54,"Excellent")'),
        ("Good Quality:", '=COUNTIF(N5:N54,"Good")'),
    ]

    row = 57
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 25
    ws.column_dimensions['N'].width = 18
    ws.column_dimensions['O'].width = 14
    ws.column_dimensions['P'].width = 35

    ws.freeze_panes = "A5"


def create_analyst_capabilities(ws, styles, validations):
    """Sheet 4: Analyst Capabilities - Skills matrix and training."""
    
    # Title
    ws.merge_cells("A1:U1")
    ws["A1"] = "Analyst Capabilities & Skills Matrix"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:U2")
    ws["A2"] = "Skills assessment for all TI analysts. Track expertise levels, certifications, and training plans."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        "Analyst_ID", "Analyst_Name", "Email", "Role", "Team", "Years_Experience_TI",
        "Years_Experience_InfoSec", "Primary_Focus_Area", "Secondary_Focus_Area",
        "Skill_OSINT", "Skill_Malware_Analysis", "Skill_MITRE_ATTACK", "Skill_Scripting_Python",
        "Skill_Threat_Modeling", "Skill_Report_Writing", "Skill_Briefing_Presentation",
        "Certifications", "Training_Planned_2025", "Training_Budget_2025",
        "Last_Performance_Review", "Capability_Gap_Identified", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 30
    
    # Data rows (20 analysts capacity: rows 5-24)
    for row in range(5, 25):
        # Analyst_ID
        cell = ws.cell(row=row, column=1)
        cell.value = f"ANALYST-{str(row-4).zfill(3)}"
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        
        # Analyst_Name, Email, Team
        for col in [2, 3, 5]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Role
        cell = ws.cell(row=row, column=4)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['analyst_role'].add(cell)
        
        # Years experience (numbers)
        for col in [6, 7]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            cell.number_format = '0'
        
        # Primary/Secondary Focus Area
        for col in [8, 9]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            validations['focus_area'].add(cell)
        
        # Skill levels (columns 10-16)
        for col in range(10, 17):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            validations['skill_level'].add(cell)
        
        # Certifications, Training Planned, Capability Gap, Notes
        for col in [17, 18, 21, 22]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Training Budget
        cell = ws.cell(row=row, column=19)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '#,##0.00'
        
        # Last Performance Review
        cell = ws.cell(row=row, column=20)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'
    
    # Apply validations
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    # Skill Level Reference Table (row 26)
    ws.merge_cells("A26:F26")
    ws["A26"] = "SKILL LEVEL REFERENCE"
    ws["A26"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A26"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    skill_ref = [
        ("Expert", "Can teach others, industry-recognized expertise"),
        ("Advanced", "Fully autonomous, handles complex scenarios"),
        ("Intermediate", "Autonomous for routine tasks, guidance for complex"),
        ("Beginner", "Requires supervision and guidance"),
        ("None", "No capability in this area"),
    ]
    
    row = 27
    for level, desc in skill_ref:
        ws[f"A{row}"] = level
        ws[f"B{row}"] = desc
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"].font = Font(name="Calibri", size=10)
        row += 1
    
    # Team Summary
    ws.merge_cells("A33:D33")
    ws["A33"] = "TEAM CAPABILITY SUMMARY"
    ws["A33"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A33"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    summary = [
        ("Total Analysts:", '=COUNTA(A5:A24)'),
        ("Avg Years TI Experience:", '=AVERAGE(F5:F24)'),
        ("Total Training Budget 2025:", '=SUM(S5:S24)'),
        ("Certifications Held:", '=COUNTA(Q5:Q24)'),
    ]
    
    row = 34
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        if "Budget" in label:
            ws[f"B{row}"].number_format = '#,##0.00'
        elif "Avg Years" in label:
            ws[f"B{row}"].number_format = '0.0'
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    for col in ['F', 'G']:
        ws.column_dimensions[col].width = 18
    for col in ['H', 'I']:
        ws.column_dimensions[col].width = 18
    for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P']:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions['Q'].width = 30
    ws.column_dimensions['R'].width = 30
    ws.column_dimensions['S'].width = 18
    ws.column_dimensions['T'].width = 18
    ws.column_dimensions['U'].width = 30
    ws.column_dimensions['V'].width = 35
    
    ws.freeze_panes = "A5"


def create_intelligence_production(ws, styles, validations):
    """Sheet 5: Intelligence Production - Track intelligence products."""
    
    # Title
    ws.merge_cells("A1:O1")
    ws["A1"] = "Intelligence Production Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:O2")
    ws["A2"] = "Track all intelligence products (reports, briefings, alerts, IOC packages). Measure quality and stakeholder feedback."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        "Product_ID", "Product_Type", "Product_Title", "Primary_Author", "Contributors",
        "Publication_Date", "TLP_Classification", "Confidence_Level", "Target_Audience",
        "Stakeholder_Feedback_Rating", "Consumption_Rate", "Actionability_Score",
        "Time_to_Produce_Hours", "Status", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 30
    
    # Data rows (60 products capacity: rows 5-64)
    for row in range(5, 65):
        # Product_ID
        cell = ws.cell(row=row, column=1)
        cell.value = f"PROD-2025-{str(row-4).zfill(3)}"
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        
        # Product_Type
        cell = ws.cell(row=row, column=2)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['product_type'].add(cell)
        
        # Product_Title, Primary_Author, Contributors, Target_Audience
        for col in [3, 4, 5, 9]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Publication_Date
        cell = ws.cell(row=row, column=6)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'
        
        # TLP_Classification
        cell = ws.cell(row=row, column=7)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['tlp'].add(cell)
        
        # Confidence_Level
        cell = ws.cell(row=row, column=8)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['confidence_level'].add(cell)
        
        # Stakeholder_Feedback_Rating (1-5)
        cell = ws.cell(row=row, column=10)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0'
        
        # Consumption_Rate (percentage)
        cell = ws.cell(row=row, column=11)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0%'
        
        # Actionability_Score (1-5)
        cell = ws.cell(row=row, column=12)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0'
        
        # Time_to_Produce_Hours
        cell = ws.cell(row=row, column=13)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0.0'
        
        # Status
        cell = ws.cell(row=row, column=14)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['product_status'].add(cell)
        
        # Notes
        cell = ws.cell(row=row, column=15)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
    
    # Apply validations
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    # Summary dashboard (row 66)
    ws.merge_cells("A66:D66")
    ws["A66"] = "PRODUCTION SUMMARY"
    ws["A66"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A66"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    summary = [
        ("Total Products:", '=COUNTA(A5:A64)'),
        ("Published This Quarter:", '=COUNTIF(N5:N64,"Published")'),
        ("Avg Feedback Rating:", '=AVERAGE(J5:J64)'),
        ("Avg Consumption Rate:", '=AVERAGE(K5:K64)'),
        ("Avg Time to Produce (hrs):", '=AVERAGE(M5:M64)'),
    ]
    
    row = 67
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        if "Rate" in label:
            ws[f"B{row}"].number_format = '0%'
        elif "Rating" in label or "hrs" in label:
            ws[f"B{row}"].number_format = '0.0'
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 22
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 30
    
    ws.freeze_panes = "A5"


# ============================================================================
# END OF SECTION 2
# ============================================================================

# ============================================================================
# SECTION 3: SHEETS 6-8 (MITRE MAPPING, QUALITY METRICS, VTL SCHEMA)
# ============================================================================

def create_mitre_mapping(ws, styles, validations):
    """Sheet 6: MITRE ATT&CK Mapping - Technique coverage analysis."""
    
    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = "MITRE ATT&CK Technique Coverage"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "Assess threat intelligence coverage of MITRE ATT&CK techniques. Focus on techniques relevant to your threat landscape."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        "Technique_ID", "Tactic", "Technique_Name", "Sub_Technique", "Priority_for_Org",
        "Coverage_Level", "Intelligence_Sources", "Last_Intelligence_Update",
        "Detection_Capability", "Use_Cases_Identified", "Gap_Analysis", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 30
    
    # Pre-populate sample critical techniques
    sample_techniques = [
        ("T1190", "Initial Access", "Exploit Public-Facing Application", "", "Critical"),
        ("T1566", "Initial Access", "Phishing", "T1566.001 - Spearphishing Attachment", "Critical"),
        ("T1059", "Execution", "Command and Scripting Interpreter", "T1059.001 - PowerShell", "High"),
        ("T1078", "Persistence", "Valid Accounts", "T1078.004 - Cloud Accounts", "High"),
        ("T1053", "Execution", "Scheduled Task/Job", "", "Medium"),
    ]
    
    row = 5
    for tech_id, tactic, tech_name, sub_tech, priority in sample_techniques:
        ws[f"A{row}"] = tech_id
        ws[f"B{row}"] = tactic
        ws[f"C{row}"] = tech_name
        ws[f"D{row}"] = sub_tech
        ws[f"E{row}"] = priority
        
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        row += 1
    
    # Additional rows (capacity 100 techniques: rows 5-104)
    for row in range(10, 105):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            
            if col == 5:  # Priority
                validations['priority'].add(cell)
            elif col == 6:  # Coverage_Level
                validations['mitre_coverage'].add(cell)
            elif col == 8:  # Last_Intelligence_Update
                cell.number_format = 'DD.MM.YYYY'
            elif col == 9:  # Detection_Capability
                validations['yes_no'].add(cell)
    
    # Apply validations
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    # Conditional formatting for coverage levels
    ws.conditional_formatting.add(
        'F5:F104',
        CellIsRule(operator='equal', formula=['"Full"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'F5:F104',
        CellIsRule(operator='equal', formula=['"None"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'F5:F104',
        CellIsRule(operator='equal', formula=['"Limited"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )
    
    # Summary dashboard (row 106)
    ws.merge_cells("A106:D106")
    ws["A106"] = "MITRE COVERAGE SUMMARY"
    ws["A106"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A106"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    summary = [
        ("Total Techniques Tracked:", '=COUNTA(A5:A104)'),
        ("Critical Priority:", '=COUNTIF(E5:E104,"Critical")'),
        ("Full Coverage:", '=COUNTIF(F5:F104,"Full")'),
        ("Partial Coverage:", '=COUNTIF(F5:F104,"Partial")'),
        ("Limited Coverage:", '=COUNTIF(F5:F104,"Limited")'),
        ("No Coverage:", '=COUNTIF(F5:F104,"None")'),
        ("With Detection Capability:", '=COUNTIF(I5:I104,"Yes")'),
    ]
    
    row = 107
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 30
    ws.column_dimensions['L'].width = 30
    
    ws.freeze_panes = "A5"


def create_quality_metrics(ws, styles, validations):
    """Sheet 7: Quality Metrics - KPIs and performance tracking."""
    
    # Title
    ws.merge_cells("A1:M1")
    ws["A1"] = "Intelligence Quality Metrics & KPIs"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:M2")
    ws["A2"] = "Track key performance indicators for intelligence collection and analysis operations."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        "KPI_ID", "KPI_Name", "Description", "Target_Value", "Current_Value",
        "Last_Month_Value", "Measurement_Period", "Performance_vs_Target",
        "Trend", "Data_Source", "Owner", "Last_Updated", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 30
    
    # Pre-populate standard KPIs
    kpis = [
        ("KPI-001", "Intelligence Products per Month", "Count of formal intelligence products", "20", "Sheet 5"),
        ("KPI-002", "Avg Time to Produce Intelligence (hrs)", "Hours from start to publication", "48", "Sheet 5"),
        ("KPI-003", "Stakeholder Feedback Rating (1-5)", "Average rating from consumers", "4.0", "Sheet 5"),
        ("KPI-004", "Intelligence Actionability Rate (%)", "% of products acted upon", "75%", "Sheet 5"),
        ("KPI-005", "Mean Time from Intelligence to Action (hrs)", "Hours from publication to action", "24", "Sheet 5"),
        ("KPI-006", "MITRE ATT&CK Coverage (%)", "% of critical techniques covered", "80%", "Sheet 6"),
        ("KPI-007", "False Positive Rate (%)", "% of incorrect/unusable intelligence", "5%", "Sheet 5"),
        ("KPI-008", "Analyst Training Completion Rate (%)", "% of planned training completed", "90%", "Sheet 4"),
        ("KPI-009", "Requirement Fulfillment Rate (%)", "% of PIRs with adequate coverage", "85%", "Sheet 2"),
        ("KPI-010", "VT Links Created per Month", "Vulnerability-threat linkages", "10", "Sheet 8"),
    ]
    
    row = 5
    for kpi_id, name, desc, target, source in kpis:
        ws[f"A{row}"] = kpi_id
        ws[f"B{row}"] = name
        ws[f"C{row}"] = desc
        ws[f"D{row}"] = target
        ws[f"J{row}"] = source
        
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            if col in [1, 8, 9]:  # ID, Performance_vs_Target, Trend (formulas)
                cell.fill = styles["formula_cell"]["fill"]
            else:
                cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Performance_vs_Target formula
        ws[f"H{row}"] = f'=IF(E{row}="","",IF(D{row}="","N/A",(E{row}/D{row}-1)*100))'
        
        # Trend formula (simplified - compare current to last month)
        ws[f"I{row}"] = f'=IF(OR(E{row}="",F{row}=""),"N/A",IF(E{row}>F{row},"Improving",IF(E{row}<F{row},"Declining","Stable")))'
        
        row += 1
    
    # Additional rows for custom KPIs (rows 15-24)
    for row in range(15, 25):
        ws[f"A{row}"] = f"KPI-{str(row-4).zfill(3)}"
        
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            if col in [1, 8, 9]:  # Formulas
                cell.fill = styles["formula_cell"]["fill"]
                if col == 8:
                    cell.value = f'=IF(E{row}="","",IF(D{row}="","N/A",(E{row}/D{row}-1)*100))'
                elif col == 9:
                    cell.value = f'=IF(OR(E{row}="",F{row}=""),"N/A",IF(E{row}>F{row},"Improving",IF(E{row}<F{row},"Declining","Stable")))'
            else:
                cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            
            if col == 12:  # Last_Updated
                cell.number_format = 'DD.MM.YYYY'
    
    # Conditional formatting for performance
    ws.conditional_formatting.add(
        'H5:H24',
        CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'H5:H24',
        CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    
    # Summary dashboard (row 26)
    ws.merge_cells("A26:D26")
    ws["A26"] = "KPI SUMMARY"
    ws["A26"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A26"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    summary = [
        ("Total KPIs Tracked:", '=COUNTA(A5:A24)'),
        ("KPIs On Target (>=100%):", '=COUNTIF(H5:H24,">=0")'),
        ("KPIs Below Target:", '=COUNTIF(H5:H24,"<0")'),
        ("Improving Trends:", '=COUNTIF(I5:I24,"Improving")'),
        ("Declining Trends:", '=COUNTIF(I5:I24,"Declining")'),
    ]
    
    row = 27
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 30
    
    ws.freeze_panes = "A5"


# ============================================================================
# SHEET 8: VULNERABILITY_LINKED_THREATS (v1.0 - WITH CVSS)
# ============================================================================

def create_vulnerability_linked_threats(ws, styles, validations):
    """
    Sheet 8: Vulnerability Linked Threats - VTL Schema v1.0 WITH CVSS
    
    CRITICAL v1.0 CHANGES:
    - Added CVSS_Version column (3): Dropdown 4.0, 3.1
    - Added CVSS_Base_Score column (4): Number 0.0-10.0
    - Added CVSS_Vector column (5): Text max 200 chars
    - Priority_Score (column 16) now AUTO-CALCULATED from CVSS
    - Added CVSS-based conditional formatting
    
    This implements the VulnerabilityThreatLink schema for bidirectional
    integration between Control 5.7 and Control 8.8.
    """
    
    # Title with CRITICAL marker
    ws.merge_cells("A1:X1")  # Expanded from T1 to X1 for 24 columns
    ws["A1"] = "WARNING: CRITICAL: Vulnerability-Linked Threats (VTL Schema v1.0 with CVSS - Control 8.8 Integration)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    
    # Instructions with CVSS note
    ws.merge_cells("A2:X2")
    ws["A2"] = "Link threat intelligence to vulnerabilities with CVSS scoring. Active exploitation + high CVSS triggers emergency patching in Control 8.8."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="FF0000", bold=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 40
    
    # Integration notice
    ws.merge_cells("A3:X3")
    ws["A3"] = "v1.0: CVSS-based priority scoring | Bidirectional: 5.7 -> 8.8 (CVSS+context) | 8.8 -> 5.7 (remediation)"
    ws["A3"].font = Font(name="Calibri", size=9, italic=True, bold=True)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Column headers (24 columns - added 3 CVSS columns)
    headers = [
        "Link_ID",
        "Vulnerability_ID",
        "CVSS_Version",           # NEW v1.0
        "CVSS_Base_Score",        # NEW v1.0
        "CVSS_Vector",            # NEW v1.0
        "Threat_Actor",
        "Threat_Actor_Type",
        "Exploitation_Status",
        "Intelligence_Source",
        "Confidence_Level",
        "Discovery_Date",
        "Last_Updated",
        "IOCs",
        "TTPs",
        "Attack_Vector",
        "Priority_Score",         # NOW AUTO-CALCULATED
        "Remediation_Urgency",
        "Business_Impact",
        "Affected_Systems_Count",
        "Critical_Assets_Affected",
        "Remediation_Status",
        "Assigned_To",
        "Related_Incidents",
        "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 35
    
    # Data rows (100 VTL records capacity: rows 5-104)
    for row in range(5, 105):
        # Link_ID (col 1 - auto-generated)
        cell = ws.cell(row=row, column=1)
        cell.value = f"VTL-20250109-{str(row-4).zfill(4)}"
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        
        # Vulnerability_ID (col 2)
        cell = ws.cell(row=row, column=2)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # CVSS_Version (col 3) - NEW v1.0
        cell = ws.cell(row=row, column=3)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['cvss_version'].add(cell)
        
        # CVSS_Base_Score (col 4) - NEW v1.0
        cell = ws.cell(row=row, column=4)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0.0'
        # Decimal validation 0.0-10.0
        dv = DataValidation(type="decimal", operator="between", 
                           formula1=0.0, formula2=10.0, allow_blank=True)
        dv.error = 'CVSS Base Score must be between 0.0 and 10.0'
        dv.errorTitle = 'Invalid CVSS Score'
        ws.add_data_validation(dv)
        dv.add(cell)
        
        # CVSS_Vector (col 5) - NEW v1.0
        cell = ws.cell(row=row, column=5)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # Threat_Actor (col 6)
        cell = ws.cell(row=row, column=6)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # Threat_Actor_Type (col 7)
        cell = ws.cell(row=row, column=7)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['threat_actor_type'].add(cell)
        
        # Exploitation_Status (col 8)
        cell = ws.cell(row=row, column=8)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['exploitation_status'].add(cell)
        
        # Intelligence_Source (col 9)
        cell = ws.cell(row=row, column=9)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # Confidence_Level (col 10)
        cell = ws.cell(row=row, column=10)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['confidence_level'].add(cell)
        
        # Discovery_Date (col 11)
        cell = ws.cell(row=row, column=11)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'
        
        # Last_Updated (col 12 - auto TODAY)
        cell = ws.cell(row=row, column=12)
        cell.value = f'=IF(B{row}="","",TODAY())'
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'
        
        # IOCs, TTPs, Attack_Vector (cols 13-15)
        for col in [13, 14, 15]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Priority_Score (col 16) - AUTO-CALCULATED v1.0
        cell = ws.cell(row=row, column=16)
        cvss_col = get_column_letter(4)
        exploit_col = get_column_letter(8)
        critical_col = get_column_letter(20)
        actor_col = get_column_letter(7)
        
        formula = f'=IF({cvss_col}{row}="","",MIN(10,{cvss_col}{row}+IF({exploit_col}{row}="Mass_Exploitation",2,IF({exploit_col}{row}="Active_Exploitation",1,0))+IF({critical_col}{row}="Yes",2,0)+IF({actor_col}{row}="Nation-State",1,0)))'
        
        cell.value = formula
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0.0'
        
        # Remediation_Urgency (col 17)
        cell = ws.cell(row=row, column=17)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['remediation_urgency'].add(cell)
        
        # Business_Impact (col 18)
        cell = ws.cell(row=row, column=18)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # Affected_Systems_Count (col 19)
        cell = ws.cell(row=row, column=19)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0'
        
        # Critical_Assets_Affected (col 20)
        cell = ws.cell(row=row, column=20)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['yes_no'].add(cell)
        
        # Remediation_Status (col 21)
        cell = ws.cell(row=row, column=21)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['remediation_status'].add(cell)
        
        # Assigned_To, Related_Incidents, Notes (cols 22-24)
        for col in [22, 23, 24]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
    
    # CVSS conditional formatting
    ws.conditional_formatting.add(
        'D5:D104',
        CellIsRule(operator='greaterThanOrEqual', formula=['9.0'],
                   fill=PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid"),
                   font=Font(color="FFFFFF", bold=True))
    )
    ws.conditional_formatting.add(
        'D5:D104',
        CellIsRule(operator='between', formula=['7.0', '8.9'],
                   fill=PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'D5:D104',
        CellIsRule(operator='between', formula=['4.0', '6.9'],
                   fill=PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'D5:D104',
        CellIsRule(operator='lessThan', formula=['4.0'],
                   fill=PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"))
    )
    
    # Exploitation status formatting
    ws.conditional_formatting.add(
        'H5:H104',
        CellIsRule(operator='equal', formula=['"Active_Exploitation"'],
                   fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                   font=Font(color="FFFFFF", bold=True))
    )
    ws.conditional_formatting.add(
        'H5:H104',
        CellIsRule(operator='equal', formula=['"Mass_Exploitation"'],
                   fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                   font=Font(color="FFFFFF", bold=True))
    )
    
    # Remediation status formatting
    ws.conditional_formatting.add(
        'U5:U104',
        CellIsRule(operator='equal', formula=['"Verified"'],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'U5:U104',
        CellIsRule(operator='equal', formula=['"Open"'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    
    # Column widths
    widths = {
        'A': 20, 'B': 16, 'C': 12, 'D': 14, 'E': 50, 'F': 20, 'G': 18, 'H': 18,
        'I': 18, 'J': 14, 'K': 12, 'L': 12, 'M': 30, 'N': 20, 'O': 30, 'P': 12,
        'Q': 16, 'R': 30, 'S': 14, 'T': 18, 'U': 16, 'V': 20, 'W': 20, 'X': 40
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_raw_intelligence_log(ws, styles, validations):
    """Sheet 4: Raw Intelligence Log - Track raw intel before analysis."""

    # Title
    ws.merge_cells("A1:P1")
    ws["A1"] = "Raw Intelligence Log"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    # Instructions
    ws.merge_cells("A2:P2")
    ws["A2"] = "Track all raw intelligence received before analysis. Log source, receipt time, initial triage, and processing status."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Column headers (Row 4)
    headers = [
        "Log_ID", "Date_Received", "Time_Received", "Source_ID", "Source_Type",
        "Intelligence_Type", "TLP_Classification", "Initial_Triage_Priority",
        "Brief_Description", "Raw_Data_Location", "Assigned_Analyst",
        "Processing_Status", "Date_Processed", "Linked_Product_ID",
        "Quality_Score", "Notes"
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()

    ws.row_dimensions[4].height = 30

    # Data rows (100 log entries capacity: rows 5-104)
    for row in range(5, 105):
        # Log_ID (auto-generated)
        cell = ws.cell(row=row, column=1)
        cell.value = f"RAW-{datetime.now().strftime('%Y')}-{str(row-4).zfill(4)}"
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()

        # Date_Received
        cell = ws.cell(row=row, column=2)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'

        # Time_Received
        cell = ws.cell(row=row, column=3)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'HH:MM'

        # Source_ID, Source_Type (text input)
        for col in [4, 5]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()

        # Intelligence_Type
        cell = ws.cell(row=row, column=6)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['intelligence_type'].add(cell)

        # TLP_Classification
        cell = ws.cell(row=row, column=7)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['tlp'].add(cell)

        # Initial_Triage_Priority
        cell = ws.cell(row=row, column=8)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['priority'].add(cell)

        # Brief_Description, Raw_Data_Location, Assigned_Analyst
        for col in [9, 10, 11]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()

        # Processing_Status
        cell = ws.cell(row=row, column=12)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['product_status'].add(cell)

        # Date_Processed
        cell = ws.cell(row=row, column=13)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'

        # Linked_Product_ID
        cell = ws.cell(row=row, column=14)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Quality_Score (1-5)
        cell = ws.cell(row=row, column=15)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

        # Notes
        cell = ws.cell(row=row, column=16)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()

    # Apply validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)

    # Conditional formatting for priority
    ws.conditional_formatting.add(
        'H5:H104',
        CellIsRule(operator='containsText', formula=['"Critical"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'H5:H104',
        CellIsRule(operator='containsText', formula=['"High"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )

    # Summary dashboard (row 106)
    ws.merge_cells("A106:D106")
    ws["A106"] = "RAW INTELLIGENCE LOG SUMMARY"
    ws["A106"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A106"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    summary_metrics = [
        ("Total Raw Intel Logged:", '=COUNTA(A5:A104)'),
        ("Critical Priority:", '=COUNTIF(H5:H104,"*Critical*")'),
        ("High Priority:", '=COUNTIF(H5:H104,"*High*")'),
        ("Pending Processing:", '=COUNTIF(L5:L104,"Draft")'),
        ("Processed:", '=COUNTIF(L5:L104,"Published")'),
        ("TLP:RED Items:", '=COUNTIF(G5:G104,"TLP:RED")'),
    ]

    row = 107
    for label, formula in summary_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 14
    ws.column_dimensions['N'].width = 18
    ws.column_dimensions['O'].width = 12
    ws.column_dimensions['P'].width = 35

    ws.freeze_panes = "A5"


def create_process_maturity(ws, styles, validations):
    """Sheet 9: Process Maturity - CMM assessment."""
    
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "Process Maturity Assessment (CMM Model)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:G2")
    ws["A2"] = "Assess capability maturity using 5-level CMM model: 1-Initial, 2-Managed, 3-Defined, 4-Quantitatively Managed, 5-Optimizing"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        "Process_Area", "Current_Level", "Target_Level", "Gap", "Evidence", "Action_Plan", "Target_Date"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 30
    
    # Pre-populate standard process areas
    process_areas = [
        "Collection Planning",
        "Analysis Methodologies",
        "Production Quality",
        "Dissemination",
        "Feedback Mechanisms",
        "Integration (Technical)",
        "Integration (Control 8.8)",
        "Staff Training",
        "Metrics & KPIs",
        "Threat Modeling",
    ]
    
    row = 5
    for area in process_areas:
        ws[f"A{row}"] = area
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].fill = styles["input_cell"]["fill"]
        ws[f"A{row}"].border = get_border()
        
        # Current_Level
        cell = ws.cell(row=row, column=2)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['maturity_level'].add(cell)
        
        # Target_Level
        cell = ws.cell(row=row, column=3)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['maturity_level'].add(cell)
        
        # Gap (formula)
        cell = ws.cell(row=row, column=4)
        cell.value = f'=IF(OR(B{row}="",C{row}=""),"N/A",VALUE(LEFT(C{row},1))-VALUE(LEFT(B{row},1)))'
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        
        # Evidence, Action_Plan
        for col in [5, 6]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Target_Date
        cell = ws.cell(row=row, column=7)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'
        
        row += 1
    
    # Apply validations
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    # Conditional formatting for gaps
    ws.conditional_formatting.add(
        'D5:D14',
        CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'D5:D14',
        CellIsRule(operator='equal', formula=['0'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    
    # Maturity Level Reference (row 16)
    ws.merge_cells("A16:G16")
    ws["A16"] = "MATURITY LEVEL REFERENCE"
    ws["A16"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A16"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    maturity_ref = [
        ("1 - Initial", "Ad hoc, reactive, inconsistent processes"),
        ("2 - Managed", "Documented processes, some repeatability"),
        ("3 - Defined", "Standardized processes, organization-wide adoption"),
        ("4 - Quantitatively Managed", "Measured and controlled processes"),
        ("5 - Optimizing", "Continuous improvement, industry-leading practices"),
    ]
    
    row = 17
    for level, desc in maturity_ref:
        ws[f"A{row}"] = level
        ws[f"B{row}"] = desc
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws.merge_cells(f"B{row}:G{row}")
        ws[f"B{row}"].font = Font(name="Calibri", size=10)
        row += 1
    
    # Summary dashboard (row 23)
    ws.merge_cells("A23:D23")
    ws["A23"] = "MATURITY SUMMARY"
    ws["A23"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A23"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    summary = [
        ("Avg Current Maturity:", '=AVERAGE(VALUE(LEFT(B5:B14,1)))'),
        ("Avg Target Maturity:", '=AVERAGE(VALUE(LEFT(C5:C14,1)))'),
        ("Total Gap (levels):", '=SUM(D5:D14)'),
        ("Areas at Target:", '=COUNTIF(D5:D14,0)'),
    ]
    
    row = 24
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"].number_format = '0.0'
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 14
    
    ws.freeze_panes = "A5"


def create_action_items(ws, styles, validations):
    """Sheet 10: Action Items - Track capability improvements."""
    
    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = "Action Items & Capability Improvements"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "Track actions to address gaps identified in collection coverage, analysis capabilities, and process maturity."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        "Action_ID", "Related_Sheet", "Issue_Description", "Priority", "Assigned_To",
        "Due_Date", "Status", "Progress_Percentage", "Completion_Date",
        "Outcome", "Lessons_Learned", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 30
    
    # Data rows (50 actions capacity: rows 5-54)
    for row in range(5, 55):
        # Action_ID
        cell = ws.cell(row=row, column=1)
        cell.value = f"ACT-572-{str(row-4).zfill(3)}"
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        
        # Related_Sheet, Issue_Description, Assigned_To, Outcome, Lessons_Learned, Notes
        for col in [2, 3, 5, 10, 11, 12]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Priority
        cell = ws.cell(row=row, column=4)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['priority'].add(cell)
        
        # Due_Date, Completion_Date
        for col in [6, 9]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            cell.number_format = 'DD.MM.YYYY'
        
        # Status
        cell = ws.cell(row=row, column=7)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['action_status'].add(cell)
        
        # Progress_Percentage
        cell = ws.cell(row=row, column=8)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0%'
    
    # Apply validations
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        'D5:D54',
        CellIsRule(operator='equal', formula=['"Critical"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'G5:G54',
        CellIsRule(operator='equal', formula=['"Closed"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'G5:G54',
        CellIsRule(operator='equal', formula=['"Blocked"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    
    # Summary dashboard (row 56)
    ws.merge_cells("A56:D56")
    ws["A56"] = "ACTION ITEMS SUMMARY"
    ws["A56"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A56"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    summary = [
        ("Total Actions:", '=COUNTA(A5:A54)'),
        ("Critical Priority:", '=COUNTIF(D5:D54,"Critical")'),
        ("High Priority:", '=COUNTIF(D5:D54,"High")'),
        ("Open:", '=COUNTIF(G5:G54,"Open")'),
        ("In Progress:", '=COUNTIF(G5:G54,"In_Progress")'),
        ("Blocked:", '=COUNTIF(G5:G54,"Blocked")'),
        ("Resolved:", '=COUNTIF(G5:G54,"Resolved")'),
        ("Closed:", '=COUNTIF(G5:G54,"Closed")'),
    ]
    
    row = 57
    for label, formula in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 35
    ws.column_dimensions['L'].width = 30
    
    ws.freeze_panes = "A5"


def create_metadata(ws, styles):
    """Sheet 11: Metadata - Workbook generation information."""
    
    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Workbook Metadata"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25
    
    generation_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    metadata = [
        ("Document ID:", "ISMS-IMP-A.5.7.2"),
        ("Document Title:", "Intelligence Collection & Analysis Assessment"),
        ("ISO 27001:2022 Control:", "A.5.7 (Threat Intelligence)"),
        ("Version:", "1.0"),
        ("Generated Date:", generation_date),
        ("Generator Script:", "generate_a57_2_collection_analysis.py"),
        ("Python Version:", "3.x"),
        ("Required Library:", "openpyxl"),
        ("", ""),
        ("Sheet Count:", "14"),
        ("Sheet 1:", "Instructions"),
        ("Sheet 2:", "Intelligence_Requirements (50 PIR/SIR/IR entries)"),
        ("Sheet 3:", "Collection_Sources (30 source entries)"),
        ("Sheet 4:", "Raw_Intelligence_Log (100 raw intel entries)"),
        ("Sheet 5:", "Intelligence_Production (60 products)"),
        ("Sheet 6:", "MITRE_Mapping (100 techniques)"),
        ("Sheet 7:", "Quality_Metrics (20 KPIs)"),
        ("Sheet 8:", "Vulnerability_Linked_Threats (50 VTL records) **CRITICAL**"),
        ("Sheet 9:", "Process_Maturity (CMM assessment)"),
        ("Sheet 10:", "Action_Items (50 actions)"),
        ("Sheet 11:", "Analysis_Tools (20 tools)"),
        ("Sheet 12:", "Threat_Actor_Profiles (30 actors)"),
        ("Sheet 13:", "Campaign_Tracking (20 campaigns)"),
        ("Sheet 14:", "Metadata (this sheet)"),
        ("", ""),
        ("VTL Integration:", "Sheet 8 implements VulnerabilityThreatLink schema"),
        ("Control 8.8 Link:", "Bidirectional data flow for vulnerability prioritization"),
        ("Emergency Patching:", "Active_Exploitation triggers emergency patch in 8.8"),
        ("", ""),
        ("Review Cycle:", "Quarterly"),
        ("Policy Reference:", "ISMS-POL-A.5.7 (All Sections)"),
        ("Related Workbooks:", "ISMS-IMP-A.5.7.1 (Sources), ISMS-IMP-A.5.7.3 (Integration)"),
        ("", ""),
        ("Notes:", "As Feynman said: 'The first principle is that you must not fool yourself'"),
        ("", "This workbook generates evidence, not theater. Use it well."),
    ]
    
    row = 3
    for label, value in metadata:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        
        if label in ["Document ID:", "Document Title:", "VTL Integration:", "Emergency Patching:"]:
            ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
        else:
            ws[f"A{row}"].font = Font(name="Calibri", size=10)
            ws[f"B{row}"].font = Font(name="Calibri", size=10)
        
        # Highlight VTL-related rows
        if "VTL" in label or "8.8" in label or "Emergency" in label:
            ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True, color="FF0000")
        
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


# ============================================================================
# MAIN EXECUTION
# ============================================================================


def create_analysis_tools(ws, styles, validations):
    """
    Sheet 11: Analysis_Tools - NEW v1.0
    Document threat intelligence analysis tools and platforms.
    """
    
    # Title
    ws.merge_cells("A1:P1")
    ws["A1"] = "Analysis Tools & Platforms (v1.0)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    # Instructions
    ws.merge_cells("A2:P2")
    ws["A2"] = "Document all threat intelligence analysis tools. Track CVSS support capability for v1.0 integration."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Column headers (16 columns)
    headers = [
        "Tool_ID", "Tool_Name", "Tool_Category", "Vendor", "License_Type",
        "Primary_Users", "Use_Cases", "Integration_Status", "Data_Sources",
        "CVSS_Support", "Last_Updated", "Version", "Training_Required",
        "Training_Status", "Cost_Annual", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 35
    
    # Data rows (50 tools capacity)
    for row in range(5, 55):
        # Tool_ID (col 1)
        cell = ws.cell(row=row, column=1)
        cell.value = f"TOOL-{str(row-4).zfill(3)}"
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        
        # Tool_Name (col 2)
        cell = ws.cell(row=row, column=2)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # Tool_Category (col 3)
        cell = ws.cell(row=row, column=3)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['tool_category'].add(cell)
        
        # Vendor, Primary_Users, Use_Cases (cols 4-7)
        for col in [4, 5, 6, 7]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # License_Type (col 5)
        cell = ws.cell(row=row, column=5)
        validations['license_type'].add(cell)
        
        # Integration_Status (col 8)
        cell = ws.cell(row=row, column=8)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['integration_status'].add(cell)
        
        # Data_Sources (col 9)
        cell = ws.cell(row=row, column=9)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # CVSS_Support (col 10) - IMPORTANT for v1.0
        cell = ws.cell(row=row, column=10)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['yes_no_na'].add(cell)
        
        # Last_Updated (col 11)
        cell = ws.cell(row=row, column=11)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'
        
        # Version (col 12)
        cell = ws.cell(row=row, column=12)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # Training_Required (col 13)
        cell = ws.cell(row=row, column=13)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['yes_no'].add(cell)
        
        # Training_Status (col 14)
        cell = ws.cell(row=row, column=14)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # Cost_Annual (col 15)
        cell = ws.cell(row=row, column=15)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '#,##0.00'
        
        # Notes (col 16)
        cell = ws.cell(row=row, column=16)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
    
    # Conditional formatting
    ws.conditional_formatting.add(
        'H5:H54',
        CellIsRule(operator='equal', formula=['"Deprecated"'],
                   fill=PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"))
    )
    
    # Column widths
    widths = {'A': 12, 'B': 25, 'C': 18, 'D': 20, 'E': 16, 'F': 20, 'G': 35, 'H': 18,
              'I': 25, 'J': 14, 'K': 12, 'L': 12, 'M': 16, 'N': 16, 'O': 12, 'P': 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width



def create_threat_actor_profiles(ws, styles, validations):
    """
    Sheet 12: Threat_Actor_Profiles - NEW v1.0
    Maintain profiles of known threat actors targeting organization or sector.
    """
    
    # Title
    ws.merge_cells("A1:T1")
    ws["A1"] = "Threat Actor Profiles (v1.0)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    # Instructions
    ws.merge_cells("A2:T2")
    ws["A2"] = "Document known threat actors. CVEs_Exploited column integrates with Sheet 8 VTL for tracking."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="8B0000", bold=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # Column headers (20 columns)
    headers = [
        "Actor_ID", "Actor_Name", "Actor_Aliases", "Actor_Type",
        "Attribution_Confidence", "Country_of_Origin", "First_Observed",
        "Last_Activity", "Targeting_Our_Sector", "Targeting_Our_Org",
        "Primary_Motivation", "Sophistication_Level", "Primary_TTPs",
        "Common_Malware", "Infrastructure_Notes", "CVEs_Exploited",
        "VTL_Records_Count", "Related_Campaigns", "Last_Update_Date", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 35
    
    # Data rows (50 actor capacity)
    for row in range(5, 55):
        # Actor_ID (col 1)
        cell = ws.cell(row=row, column=1)
        cell.value = f"ACTOR-{str(row-4).zfill(3)}"
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        
        # Actor_Name, Actor_Aliases (cols 2-3)
        for col in [2, 3]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Actor_Type (col 4)
        cell = ws.cell(row=row, column=4)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['threat_actor_type'].add(cell)
        
        # Attribution_Confidence (col 5)
        cell = ws.cell(row=row, column=5)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['confidence_level'].add(cell)
        
        # Country_of_Origin (col 6)
        cell = ws.cell(row=row, column=6)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # First_Observed, Last_Activity (cols 7-8)
        for col in [7, 8]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            cell.number_format = 'DD.MM.YYYY'
        
        # Targeting_Our_Sector (col 9)
        cell = ws.cell(row=row, column=9)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['yes_no_unknown'].add(cell)
        
        # Targeting_Our_Org (col 10)
        cell = ws.cell(row=row, column=10)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['targeting_status'].add(cell)
        
        # Primary_Motivation (col 11)
        cell = ws.cell(row=row, column=11)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['motivation'].add(cell)
        
        # Sophistication_Level (col 12)
        cell = ws.cell(row=row, column=12)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['sophistication'].add(cell)
        
        # Primary_TTPs, Common_Malware, Infrastructure_Notes (cols 13-15)
        for col in [13, 14, 15]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # CVEs_Exploited (col 16) - INTEGRATION with Sheet 8
        cell = ws.cell(row=row, column=16)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # VTL_Records_Count (col 17)
        cell = ws.cell(row=row, column=17)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0'
        
        # Related_Campaigns (col 18)
        cell = ws.cell(row=row, column=18)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # Last_Update_Date (col 19)
        cell = ws.cell(row=row, column=19)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'
        
        # Notes (col 20)
        cell = ws.cell(row=row, column=20)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
    
    # Conditional formatting
    ws.conditional_formatting.add(
        'J5:J54',
        CellIsRule(operator='equal', formula=['"Confirmed"'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'J5:J54',
        CellIsRule(operator='equal', formula=['"Suspected"'],
                   fill=PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"))
    )
    
    # Column widths
    widths = {'A': 12, 'B': 25, 'C': 25, 'D': 18, 'E': 18, 'F': 18, 'G': 12, 'H': 12,
              'I': 18, 'J': 16, 'K': 18, 'L': 16, 'M': 30, 'N': 25, 'O': 35, 'P': 30,
              'Q': 14, 'R': 25, 'S': 12, 'T': 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width



def create_campaign_tracking(ws, styles, validations):
    """
    Sheet 13: Campaign_Tracking - NEW v1.0
    Track and analyze threat campaigns with VLOOKUP integration to Sheet 12.
    """
    
    # Title
    ws.merge_cells("A1:X1")
    ws["A1"] = "Campaign Tracking (v1.0 - VLOOKUP + CVSS Integration)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    # Instructions
    ws.merge_cells("A2:X2")
    ws["A2"] = "Track threat campaigns. Actor_Name auto-populated via VLOOKUP from Sheet 12. CVEs_CVSS_Max references Sheet 8."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="8B0000", bold=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 35
    
    # Column headers (24 columns)
    headers = [
        "Campaign_ID", "Campaign_Name", "Actor_ID", "Actor_Name",
        "Campaign_Start_Date", "Campaign_End_Date", "Campaign_Status",
        "Target_Sectors", "Target_Geographies", "Our_Sector_Targeted",
        "Our_Org_Targeted", "Primary_Objective", "Attack_Vectors",
        "TTPs_Used", "CVEs_Exploited", "CVEs_CVSS_Max",
        "IOCs_Count", "VTL_Records_Created", "Incidents_Our_Org",
        "Intelligence_Sources", "Threat_Level", "Monitoring_Status",
        "Last_Update_Date", "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header.replace("_", " ")
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = get_border()
    
    ws.row_dimensions[4].height = 35
    
    # Data rows (50 campaign capacity)
    for row in range(5, 55):
        # Campaign_ID (col 1)
        cell = ws.cell(row=row, column=1)
        cell.value = f"CAMP-2025-{str(row-4).zfill(3)}"
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        
        # Campaign_Name (col 2)
        cell = ws.cell(row=row, column=2)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # Actor_ID (col 3) - dropdown references Sheet 12
        cell = ws.cell(row=row, column=3)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # Actor_Name (col 4) - VLOOKUP from Sheet 12
        cell = ws.cell(row=row, column=4)
        formula = f'=IFERROR(VLOOKUP(C{row},Threat_Actor_Profiles!A:B,2,FALSE),"Unknown")'
        cell.value = formula
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        
        # Campaign_Start_Date, Campaign_End_Date (cols 5-6)
        for col in [5, 6]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            cell.number_format = 'DD.MM.YYYY'
        
        # Campaign_Status (col 7)
        cell = ws.cell(row=row, column=7)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['campaign_status'].add(cell)
        
        # Target_Sectors, Target_Geographies (cols 8-9)
        for col in [8, 9]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # Our_Sector_Targeted (col 10)
        cell = ws.cell(row=row, column=10)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['yes_no_unknown'].add(cell)
        
        # Our_Org_Targeted (col 11)
        cell = ws.cell(row=row, column=11)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['targeting_status'].add(cell)
        
        # Primary_Objective (col 12)
        cell = ws.cell(row=row, column=12)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['campaign_objective'].add(cell)
        
        # Attack_Vectors, TTPs_Used, CVEs_Exploited (cols 13-15)
        for col in [13, 14, 15]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
        
        # CVEs_CVSS_Max (col 16) - References Sheet 8 CVSS data
        cell = ws.cell(row=row, column=16)
        cell.fill = styles["formula_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = '0.0'
        # Note: Complex formula - simplified placeholder
        cell.value = ""  # User to populate based on Sheet 8 data
        
        # IOCs_Count, VTL_Records_Created, Incidents_Our_Org (cols 17-19)
        for col in [17, 18, 19]:
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = get_border()
            cell.number_format = '0'
        
        # Intelligence_Sources (col 20)
        cell = ws.cell(row=row, column=20)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        
        # Threat_Level (col 21)
        cell = ws.cell(row=row, column=21)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['threat_level'].add(cell)
        
        # Monitoring_Status (col 22)
        cell = ws.cell(row=row, column=22)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        validations['monitoring_status'].add(cell)
        
        # Last_Update_Date (col 23)
        cell = ws.cell(row=row, column=23)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
        cell.number_format = 'DD.MM.YYYY'
        
        # Notes (col 24)
        cell = ws.cell(row=row, column=24)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = get_border()
    
    # Conditional formatting
    ws.conditional_formatting.add(
        'K5:K54',
        CellIsRule(operator='equal', formula=['"Confirmed"'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'U5:U54',
        CellIsRule(operator='equal', formula=['"Critical"'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                   font=Font(bold=True))
    )
    
    # Column widths
    widths = {'A': 16, 'B': 30, 'C': 12, 'D': 25, 'E': 12, 'F': 12, 'G': 14, 'H': 30,
              'I': 25, 'J': 16, 'K': 16, 'L': 20, 'M': 30, 'N': 30, 'O': 30, 'P': 12,
              'Q': 12, 'R': 14, 'S': 14, 'T': 25, 'U': 14, 'V': 16, 'W': 12, 'X': 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def main():
    """Main generation function for v1.0 (14 sheets - aligned with IMP specification)."""
    logger.info("")
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.5.7.2 v1.0 - Intelligence Collection & Analysis Generator")
    logger.info("ISO/IEC 27001:2022 Control A.5.7 (Threat Intelligence)")
    logger.info("=" * 80)
    logger.info("")

    wb = create_workbook()
    styles = setup_styles()
    validations = setup_validations()

    logger.info("Generating sheets...")

    # Sheet 1: Instructions
    logger.info("  [1/14] Instructions")
    ws = wb["Instructions"]
    create_instructions(ws, styles)

    # Sheet 2: Intelligence_Requirements (PIR/SIR/IR per IMP spec)
    logger.info("  [2/14] Intelligence_Requirements")
    ws = wb["Intelligence_Requirements"]
    create_intelligence_requirements(ws, styles, validations)

    # Sheet 3: Collection_Sources
    logger.info("  [3/14] Collection_Sources")
    ws = wb["Collection_Sources"]
    create_collection_sources(ws, styles, validations)

    # Sheet 4: Raw_Intelligence_Log
    logger.info("  [4/14] Raw_Intelligence_Log")
    ws = wb["Raw_Intelligence_Log"]
    create_raw_intelligence_log(ws, styles, validations)

    # Sheet 5: Intelligence_Production
    logger.info("  [5/14] Intelligence_Production")
    ws = wb["Intelligence_Production"]
    create_intelligence_production(ws, styles, validations)

    # Sheet 6: MITRE_Mapping
    logger.info("  [6/14] MITRE_Mapping")
    ws = wb["MITRE_Mapping"]
    create_mitre_mapping(ws, styles, validations)

    # Sheet 7: Quality_Metrics
    logger.info("  [7/14] Quality_Metrics")
    ws = wb["Quality_Metrics"]
    create_quality_metrics(ws, styles, validations)

    # Sheet 8: Vulnerability_Linked_Threats (v1.0 WITH CVSS)
    logger.info("  [8/14] Vulnerability_Linked_Threats (v1.0 - CVSS INTEGRATED)")
    ws = wb["Vulnerability_Linked_Threats"]
    create_vulnerability_linked_threats(ws, styles, validations)

    # Sheet 9: Process_Maturity
    logger.info("  [9/14] Process_Maturity")
    ws = wb["Process_Maturity"]
    create_process_maturity(ws, styles, validations)

    # Sheet 10: Action_Items
    logger.info("  [10/14] Action_Items")
    ws = wb["Action_Items"]
    create_action_items(ws, styles, validations)

    # Sheet 11: Analysis_Tools
    logger.info("  [11/14] Analysis_Tools")
    ws = wb["Analysis_Tools"]
    create_analysis_tools(ws, styles, validations)

    # Sheet 12: Threat_Actor_Profiles
    logger.info("  [12/14] Threat_Actor_Profiles")
    ws = wb["Threat_Actor_Profiles"]
    create_threat_actor_profiles(ws, styles, validations)

    # Sheet 13: Campaign_Tracking
    logger.info("  [13/14] Campaign_Tracking")
    ws = wb["Campaign_Tracking"]
    create_campaign_tracking(ws, styles, validations)

    # Sheet 14: Metadata
    logger.info("  [14/14] Metadata")
    ws = wb["Metadata"]
    create_metadata(ws, styles)

    # Save workbook
    filename = f"ISMS-IMP-A.5.7.2_Collection_Analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    logger.info("")
    logger.info("=" * 80)
    logger.info("SUCCESS!")
    logger.info("=" * 80)
    logger.info("")
    logger.info(f"Generated: {filename}")
    logger.info("")
    logger.info("Workbook Contents (v1.0 - IMP Aligned):")
    logger.info("  - 14 sheets (aligned with IMP specification)")
    logger.info("  - Intelligence_Requirements: 50 PIR/SIR/IR entries")
    logger.info("  - Collection_Sources: 50 source records")
    logger.info("  - Raw_Intelligence_Log: 100 raw intel entries")
    logger.info("  - Intelligence_Production: 100 product records")
    logger.info("  - MITRE_Mapping: ATT&CK technique coverage")
    logger.info("  - Quality_Metrics: KPI dashboard")
    logger.info("  - Vulnerability_Linked_Threats: 100 VTL records WITH CVSS v4.0/3.1")
    logger.info("  - Process_Maturity: CMM capability assessment")
    logger.info("  - Action_Items: 100 improvement actions")
    logger.info("  - Analysis_Tools: 50 tool records")
    logger.info("  - Threat_Actor_Profiles: 50 actor profiles")
    logger.info("  - Campaign_Tracking: 50 campaign records")
    logger.info("  - Metadata: Assessment metadata and approvals")
    logger.info("")
    logger.info("CRITICAL v1.0 FEATURES:")
    logger.info("  Sheet 8 (Vulnerability_Linked_Threats) WITH CVSS:")
    logger.info("    - CVSS_Version column (4.0, 3.1)")
    logger.info("    - CVSS_Base_Score column (0.0-10.0)")
    logger.info("    - CVSS_Vector column (max 200 chars)")
    logger.info("    - Priority_Score AUTO-CALCULATED from CVSS + threat factors")
    logger.info("    - CVSS-based conditional formatting (Critical/High/Medium/Low)")
    logger.info("    - Emergency highlighting for Mass Exploitation")
    logger.info("")
    logger.info("  Bidirectional data flow with Control 8.8:")
    logger.info("    5.7 -> 8.8: CVSS scores + threat context")
    logger.info("    8.8 -> 5.7: Remediation status updates")
    logger.info("    Active exploitation + high CVSS -> Emergency patching trigger")
    logger.info("")
    logger.info("Next Steps:")
    logger.info("  1. Run sanity check:")
    logger.info("     python3 excel_sanity_check_a57_2.py " + filename)
    logger.info("  2. Complete workbook with organizational data")
    logger.info("  3. Link to Control 8.8 vulnerability workbooks")
    logger.info("")
    logger.info("=" * 80)


if __name__ == "__main__":
    main()


# ============================================================================
# END OF GENERATE_A57_2_COLLECTION_ANALYSIS.PY
# ============================================================================
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - IMP ALIGNMENT COMPLETE
# QA_TOOL: Claude Code Standardization
# CHANGES: Aligned sheets with IMP specification:
#   - Renamed create_collection_coverage -> create_intelligence_requirements (Sheet 2)
#   - Renamed create_analysis_framework -> create_collection_sources (Sheet 3)
#   - Sheet 4: Raw_Intelligence_Log (uses create_raw_intelligence_log)
#   - Sheet 9: Process_Maturity (uses create_process_maturity, replaces Analyst_Capabilities)
#   - Updated Instructions sheet guide and completion instructions
#   - Updated Metadata sheet references
# =============================================================================
