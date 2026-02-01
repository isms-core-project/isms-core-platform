#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.5-6.1 - Authority Contacts Register
================================================================================

ISO/IEC 27001:2022 Controls A.5.5 & A.5.6: External Communications
Assessment Domain 1 of 4: Authority Contacts Register

This script generates a comprehensive Excel workbook for managing contacts with
authorities including law enforcement, regulatory bodies, and government agencies
as required by ISO 27001:2022 Control A.5.5.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.5-6.1"
WORKBOOK_NAME = "Authority Contacts Register"
CONTROL_ID = "A.5.5-6"
CONTROL_NAME = "Contact with Authorities & Special Interest Groups"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls A.5.5 & A.5.6: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions sheet."""
    ws.title = "Instructions"

    instructions = [
        ["ISMS-IMP-A.5.5-6.1 - Authority Contacts Register"],
        [""],
        ["PURPOSE"],
        ["This workbook maintains a register of contacts with authorities including law enforcement,"],
        ["regulatory bodies, government agencies, and emergency services as required by ISO 27001:2022 Control A.5.5."],
        [""],
        ["ISO 27001:2022 CONTROL A.5.5 - CONTACT WITH AUTHORITIES"],
        ["Appropriate contacts with relevant authorities should be maintained."],
        [""],
        ["SHEETS"],
        ["1. Instructions - This guidance sheet"],
        ["2. Authority_Registry - Master list of authority contacts"],
        ["3. Contact_Types - Authority categories and escalation paths"],
        ["4. Communication_Log - Record of communications with authorities"],
        ["5. Verification_Register - Annual verification of contact details"],
        ["6. Evidence_Register - Evidence documentation for audits"],
        ["7. Approval_SignOff - Management approval workflow"],
        [""],
        ["WORKFLOW"],
        ["1. Identify all relevant authorities (regulatory, law enforcement, emergency)"],
        ["2. Document contact details and escalation paths"],
        ["3. Assign internal owners responsible for each authority relationship"],
        ["4. Verify contact details annually (minimum)"],
        ["5. Log all significant communications"],
        ["6. Review and update register quarterly"],
        [""],
        ["AUTHORITY CATEGORIES"],
        ["- Law Enforcement (Police, Cybercrime units, Interpol)"],
        ["- Data Protection Authorities (FDPIC, ICO, CNIL, etc.)"],
        ["- Financial Regulators (FINMA, BaFin, SEC, etc.)"],
        ["- Sector Regulators (industry-specific bodies)"],
        ["- Emergency Services (Fire, Ambulance, Utilities)"],
        ["- Government Cybersecurity (NCSC, CISA, ENISA, etc.)"],
        ["- Standards Bodies (ISO National Members)"],
        [""],
        ["DATA VALIDATION"],
        ["- Yellow cells are input fields"],
        ["- Dropdown lists enforce valid values"],
        ["- Contact_ID must be unique (AUTH-XXXX format)"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SHEETS", "WORKFLOW", "AUTHORITY CATEGORIES", "DATA VALIDATION",
                          "ISO 27001:2022 CONTROL A.5.5 - CONTACT WITH AUTHORITIES"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 90


def create_authority_registry_sheet(ws):
    """Create the Authority Registry sheet."""
    ws.title = "Authority_Registry"

    headers = [
        "Contact_ID", "Authority_Name", "Authority_Type", "Jurisdiction",
        "Primary_Contact_Name", "Primary_Contact_Title", "Primary_Phone",
        "Primary_Email", "Secondary_Contact_Name", "Secondary_Phone",
        "Website", "Physical_Address", "Internal_Owner", "Owner_Department",
        "Relationship_Status", "Last_Contact_Date", "Next_Review_Date",
        "Escalation_Trigger", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validation dropdowns
    type_dv = DataValidation(
        type="list",
        formula1='"Law Enforcement,Data Protection Authority,Financial Regulator,Sector Regulator,Emergency Services,Government Cybersecurity,Standards Body,Other"'
    )
    ws.add_data_validation(type_dv)
    type_dv.add('C2:C100')

    jurisdiction_dv = DataValidation(
        type="list",
        formula1='"Switzerland,Germany,Austria,EU,United Kingdom,United States,International,Other"'
    )
    ws.add_data_validation(jurisdiction_dv)
    jurisdiction_dv.add('D2:D100')

    status_dv = DataValidation(type="list", formula1='"Active,Inactive,Under Review,Pending"')
    ws.add_data_validation(status_dv)
    status_dv.add('O2:O100')

    # Format input rows
    for row in range(2, 31):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 35, 25, 15, 25, 25, 18, 30, 25, 18, 35, 40, 20, 18, 15, 15, 15, 30, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_contact_types_sheet(ws):
    """Create the Contact Types sheet."""
    ws.title = "Contact_Types"

    headers = [
        "Type_Code", "Authority_Type", "Description", "Typical_Scenarios",
        "Mandatory_Contact", "Response_SLA", "Escalation_Path",
        "Required_Information", "Example_Authorities"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate with standard authority types
    authority_types = [
        ("LE", "Law Enforcement", "Police and cybercrime investigation units",
         "Security incidents, theft, criminal activity", "Yes (for crimes)",
         "Immediate", "CISO > Legal > CEO", "Incident details, evidence preservation",
         "Cantonal Police, Fedpol, Europol"),
        ("DPA", "Data Protection Authority", "Personal data protection regulators",
         "Data breaches affecting individuals", "Yes (72h breach notification)",
         "72 hours", "DPO > Legal > CISO", "Breach scope, affected individuals, mitigation",
         "FDPIC (CH), ICO (UK), CNIL (FR)"),
        ("FIN", "Financial Regulator", "Financial services regulators",
         "Financial system impacts, regulatory reporting", "Conditional",
         "Per regulation", "CFO > Legal > CEO", "Financial impact, regulatory requirements",
         "FINMA (CH), BaFin (DE), SEC (US)"),
        ("SEC", "Sector Regulator", "Industry-specific regulatory bodies",
         "Industry compliance, sector-specific incidents", "Conditional",
         "Per regulation", "Compliance > Legal", "Sector-specific details",
         "OFCOM, BAFU, Swissmedic"),
        ("EME", "Emergency Services", "Fire, ambulance, utilities",
         "Physical emergencies, facility incidents", "Yes (emergencies)",
         "Immediate", "Facility Mgr > HR > CEO", "Location, nature of emergency",
         "Fire Brigade, Ambulance, Police"),
        ("CYB", "Government Cybersecurity", "National cybersecurity centers",
         "Cyber threats, national security concerns", "Recommended",
         "24-48 hours", "CISO > Legal", "Threat details, IOCs, impact",
         "NCSC (CH), BSI (DE), CISA (US)"),
        ("STD", "Standards Body", "Certification and standards organizations",
         "Certification matters, standard interpretations", "No",
         "Per agreement", "Quality > Compliance", "Certification scope, queries",
         "SNV, ISO National Members"),
    ]

    row = 2
    for type_data in authority_types:
        for col, value in enumerate(type_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Column widths
    widths = [10, 25, 40, 35, 18, 15, 25, 35, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_communication_log_sheet(ws):
    """Create the Communication Log sheet."""
    ws.title = "Communication_Log"

    headers = [
        "Log_ID", "Contact_ID", "Authority_Name", "Communication_Date",
        "Communication_Type", "Direction", "Subject", "Summary",
        "Our_Representative", "Authority_Representative", "Outcome",
        "Follow_Up_Required", "Follow_Up_Date", "Evidence_Reference", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(
        type="list",
        formula1='"Meeting,Phone Call,Email,Letter,Report Submission,Notification,Other"'
    )
    ws.add_data_validation(type_dv)
    type_dv.add('E2:E200')

    direction_dv = DataValidation(type="list", formula1='"Outbound,Inbound"')
    ws.add_data_validation(direction_dv)
    direction_dv.add('F2:F200')

    followup_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(followup_dv)
    followup_dv.add('L2:L200')

    # Format input rows
    for row in range(2, 51):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 12, 30, 15, 18, 12, 35, 50, 20, 25, 25, 15, 15, 25, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_verification_register_sheet(ws):
    """Create the Verification Register sheet."""
    ws.title = "Verification_Register"

    headers = [
        "Verification_ID", "Contact_ID", "Authority_Name", "Verification_Date",
        "Verified_By", "Verification_Method", "Contact_Details_Correct",
        "Escalation_Path_Tested", "Website_Accessible", "Relationship_Active",
        "Issues_Found", "Actions_Required", "Next_Verification", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    method_dv = DataValidation(
        type="list",
        formula1='"Phone Test,Email Test,Website Check,Meeting,Document Review"'
    )
    ws.add_data_validation(method_dv)
    method_dv.add('F2:F100')

    yn_dv = DataValidation(type="list", formula1='"Yes,No,N/A"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('G2:G100')
    yn_dv.add('H2:H100')
    yn_dv.add('I2:I100')
    yn_dv.add('J2:J100')

    # Format input rows
    for row in range(2, 31):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [15, 12, 30, 15, 20, 20, 20, 20, 18, 18, 30, 30, 18, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_evidence_register_sheet(ws):
    """Create the Evidence Register sheet."""
    ws.title = "Evidence_Register"

    headers = [
        "Evidence_ID", "Evidence_Type", "Description", "Related_Contact_ID",
        "Date_Created", "Created_By", "Storage_Location",
        "Retention_Period", "Review_Date", "Status", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(
        type="list",
        formula1='"Contact List,Communication Record,Verification Evidence,MOU/Agreement,Notification Receipt,Meeting Minutes,Other"'
    )
    ws.add_data_validation(type_dv)
    type_dv.add('B2:B100')

    status_dv = DataValidation(type="list", formula1='"Current,Archived,Pending Review"')
    ws.add_data_validation(status_dv)
    status_dv.add('J2:J100')

    # Format input rows
    for row in range(2, 31):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 22, 40, 18, 15, 20, 40, 15, 15, 15, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_approval_signoff_sheet(ws):
    """Create the Approval Sign-Off sheet."""
    ws.title = "Approval_SignOff"

    headers = [
        "Approval_ID", "Review_Period", "Review_Date", "Reviewer_Name",
        "Reviewer_Role", "Registry_Complete", "Contacts_Verified",
        "Communication_Log_Current", "Approval_Status", "Signature_Date",
        "Next_Review_Date", "Comments"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    role_dv = DataValidation(
        type="list",
        formula1='"CISO,Legal Counsel,DPO,Compliance Officer,Risk Manager,CEO"'
    )
    ws.add_data_validation(role_dv)
    role_dv.add('E2:E20')

    yn_dv = DataValidation(type="list", formula1='"Yes,No,Partial"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('F2:F20')
    yn_dv.add('G2:G20')
    yn_dv.add('H2:H20')

    status_dv = DataValidation(type="list", formula1='"Approved,Rejected,Pending,Conditional"')
    ws.add_data_validation(status_dv)
    status_dv.add('I2:I20')

    # Format input rows
    for row in range(2, 11):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 15, 15, 25, 20, 18, 18, 22, 15, 15, 18, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_workbook():
    """Generate the complete workbook."""
    logger.info("=" * 70)
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")
    logger.info("=" * 70)

    wb = Workbook()

    # Create all sheets
    create_instructions_sheet(wb.active)
    create_authority_registry_sheet(wb.create_sheet())
    create_contact_types_sheet(wb.create_sheet())
    create_communication_log_sheet(wb.create_sheet())
    create_verification_register_sheet(wb.create_sheet())
    create_evidence_register_sheet(wb.create_sheet())
    create_approval_signoff_sheet(wb.create_sheet())

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")
    logger.info("=" * 70)

    return OUTPUT_FILENAME


if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.5-6 External Communications control
# =============================================================================
