#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.32.2 - Change Types & Categories Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.32: Change Management
Assessment Domain 2 of 4: Change Classification & Risk Management

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific change classification schemes and assessment
requirements.

Key customization areas:
1. Standard change catalog (match your pre-approved changes)
2. Change risk classification criteria (adapt to your risk appetite)
3. Emergency change triggers (customize to your business context)
4. Change calendar blackout periods (align with your business cycles)
5. Approval workflows per change type (match your governance)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.32 Change Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
change classification, risk assessment, and category management against
ISO 27001:2022 Control A.8.32 requirements.

**Purpose:**
Enables systematic assessment of change type definitions, standard change
catalogs, normal change procedures, emergency change processes, risk
classification matrices, and change calendar management.

**Assessment Scope:**
- Standard change catalog and pre-authorization criteria
- Normal change risk assessment and CAB review procedures
- Emergency change triggers and fast-track approvals
- Change risk classification (Impact × Likelihood matrices)
- Change calendar management and blackout periods
- Change type compliance verification
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and classification standards
2. Standard_Changes_Catalog - Pre-approved low-risk changes
3. Normal_Changes_Assessment - CAB-reviewed changes and procedures
4. Emergency_Changes - Fast-track emergency change process
5. Change_Risk_Classification - Risk matrix and scoring
6. Change_Calendar_Management - Blackout periods and scheduling
7. Summary_Dashboard - Compliance metrics and analytics
8. Evidence_Register - Audit evidence tracking
9. Approval_Sign_Off - Stakeholder approval workflow

**Key Features:**
- Technology-agnostic assessment (works with any ITSM tool)
- Standard change catalog template with customization guidance
- Risk matrix with Impact × Likelihood methodology
- Change calendar integration for blackout period management
- Automated compliance calculations
- Evidence linkage for audit traceability
- Integration with A.8.32.5 Compliance Dashboard

**Integration:**
This assessment feeds into the A.8.32.5 Compliance Dashboard, which
consolidates data from all four change management assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a832_2_change_types.py

Output:
    File: ISMS_A_8_32_2_Change_Types_Categories_Assessment_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Define your organization's standard change catalog
    2. Document normal change risk assessment procedures
    3. Define emergency change triggers specific to your business
    4. Customize risk matrix to your risk appetite
    5. Document change calendar blackout periods
    6. Review Summary_Dashboard for compliance metrics
    7. Collect and link audit evidence
    8. Feed results into A.8.32.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.32
Assessment Domain:    2 of 4 (Change Classification & Risk Management)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.32: Change Management Policy (Governance)
    - ISMS-IMP-A.8.32.2: Change Types & Categories Implementation Guide
    - ISMS-IMP-A.8.32.1: Change Process Assessment (Domain 1)
    - ISMS-IMP-A.8.32.3: Environment Separation Assessment (Domain 3)
    - ISMS-IMP-A.8.32.4: Testing & Validation Assessment (Domain 4)
    - ISMS-IMP-A.8.32.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.32.2 specification
    - Supports comprehensive change classification evaluation
    - Integrated with A.8.32.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Standard Change Philosophy:**
Standard changes are pre-approved, low-risk, well-understood changes that follow
a documented procedure. They don't require CAB approval but DO require logging
and monitoring for compliance.

**Risk Classification Flexibility:**
Customize the risk matrix to your organization's risk appetite. A startup might
accept higher risks than a regulated financial institution. The framework is
intentionally flexible.

**Emergency Change Reality:**
Emergency changes will happen. The goal is controlled urgency, not prevention.
Document clear triggers, fast-track approvals, and mandatory post-implementation
reviews.

**Audit Considerations:**
Auditors will verify that change types are clearly defined, risk assessments
are systematic, and emergency changes have proper justification and review.

**Data Protection:**
Assessment workbooks contain sensitive information about change management
practices, risk tolerances, and operational procedures. Handle according to
your data classification policies.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.32.2"
WORKBOOK_NAME = "Change Types & Categories Assessment"
CONTROL_ID = "A.8.32"
CONTROL_NAME = "Change Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
CLOCK = '\U0001F552'  # 🕒 Clock
WRENCH = '\U0001F527' # 🔧 Wrench
ROCKET = '\U0001F680' # 🚀 Rocket
GEAR = '\u2699'       # ⚙️  Gear
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP-A.8.32.2 spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches ISMS-IMP-A.8.32.2 specification (10 sheets)
    sheets = [
        "Instructions & Legend",
        "Standard_Changes_Catalog",
        "Normal_Change_Classification",
        "Emergency_Change_Procedures",
        "Risk_Assessment_Matrix",
        "Change_Calendar_Management",
        "Classification_Metrics",
        "Evidence_Register",
        "Summary_Dashboard",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "calculated_cell": {
            "fill": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "risk_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "risk_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "risk_high": {
            "fill": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        },
        "risk_critical": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell. Creates NEW style objects."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & DROPDOWN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    CUSTOMIZE: Adapt dropdown values for change classification context.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'definition_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Defined,⚠️ Partial,❌ Not Defined,📋 Planned,N/A"',
            allow_blank=False
        ),
        'change_category': DataValidation(
            type="list",
            formula1='"Infrastructure,Application,Security,Data,Network,Cloud,User Access,Configuration,Other"',
            allow_blank=False
        ),
        'frequency': DataValidation(
            type="list",
            formula1='"Daily,Weekly,Monthly,Quarterly,Annual,On-Demand,Rare"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Low,Medium,High,Critical"',
            allow_blank=False
        ),
        'std_change_risk': DataValidation(
            type="list",
            formula1='"Low,Medium"',
            allow_blank=False
        ),
        'std_change_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Active,⚠️ Under Review,❌ Retired,📋 Proposed"',
            allow_blank=False
        ),
        'cab_requirement': DataValidation(
            type="list",
            formula1='"Mandatory,Recommended,Optional,Not Required"',
            allow_blank=False
        ),
        'approval_authority': DataValidation(
            type="list",
            formula1='"Change Manager,CAB,Service Owner,CISO,CIO,Multiple Required,Other"',
            allow_blank=False
        ),
        'implementation_level': DataValidation(
            type="list",
            formula1=f'"{CHECK} Always,⚠️ Sometimes,❌ Never,📋 Planned"',
            allow_blank=False
        ),
        'exceptions_allowed': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes,❌ No,⚠️ Case-by-Case,N/A"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Yes,⚠️ Partial,❌ No"',
            allow_blank=False
        ),
        'emergency_criteria_definition': DataValidation(
            type="list",
            formula1=f'"{CHECK} Clearly Defined,⚠️ Partially,❌ Not Defined"',
            allow_blank=False
        ),
        'user_count': DataValidation(
            type="list",
            formula1='"<10 users,10-50,50-100,100-1000,>1000 users,All users,Business-critical,Other"',
            allow_blank=False
        ),
        'downtime_potential': DataValidation(
            type="list",
            formula1='"<15 min,15-60 min,1-2 hours,2-8 hours,>8 hours,Irreversible,Business-critical,Other"',
            allow_blank=False
        ),
        'data_loss_risk': DataValidation(
            type="list",
            formula1='"None,Minimal,Moderate,High"',
            allow_blank=False
        ),
        'complexity': DataValidation(
            type="list",
            formula1='"Simple,Moderate,Complex,Very Complex"',
            allow_blank=False
        ),
        'testing_maturity': DataValidation(
            type="list",
            formula1='"Extensively Tested,Well Tested,Limited Testing,Untested"',
            allow_blank=False
        ),
        'convening_method': DataValidation(
            type="list",
            formula1='"Conference Call,Video,IM Group,Emergency Hotline,Other"',
            allow_blank=False
        ),
        'applicable_changes': DataValidation(
            type="list",
            formula1='"All,Standard Only,Normal Only,Emergency Only,Custom"',
            allow_blank=False
        ),
        'blackout_reason': DataValidation(
            type="list",
            formula1='"Financial Close,Holiday Period,Peak Business,Audit,Other"',
            allow_blank=False
        ),
        'conflict_detection': DataValidation(
            type="list",
            formula1='"Automated,Manual,None"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Catalog,Procedure,Criteria Doc,Risk Matrix,Calendar,Meeting Minutes,Metrics Report,Approval Record,Training Material,Other"',
            allow_blank=False
        ),
        'verification_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Verified,⚠️ Pending,❌ Not Verified"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Final,⚠️ Requires Remediation,📋 Draft,❌ Re-assessment Required"',
            allow_blank=False
        ),
        'review_recommendation': DataValidation(
            type="list",
            formula1=f'"{CHECK} Approve,⚠️ Approve with Conditions,❌ Reject,📋 Require Rework"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1=f'"{CHECK} Approved,⚠️ Approved with Conditions,❌ Rejected"',
            allow_blank=False
        ),
        'regulatory_review': DataValidation(
            type="list",
            formula1='"Yes,No,To Be Determined"',
            allow_blank=False
        ),
    }

    # Add all validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)

    return validations


# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet with usage guidance."""
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "ISMS-IMP-A.8.32.2 – Change Types & Categories Assessment"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "ISO/IEC 27001:2022 - Control A.8.32: Change Management"
    apply_style(ws["A2"], styles["subheader"])

    # Document Information Block
    row = 4
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    
    doc_info = [
        ("Document ID:", "ISMS-IMP-A.8.32.2"),
        ("Assessment Area:", "Change Types & Categories"),
        ("Related Policy:", "ISMS-POL-A.8.32-S2.2"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[Organization]"),
        ("Review Cycle:", "Quarterly"),
    ]
    
    row += 1
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if "USER INPUT" in value:
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    # How to Use This Workbook
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    instructions = [
        "1. Complete the Standard_Changes_Catalog with YOUR pre-approved changes",
        "2. Document YOUR normal change assessment criteria",
        "3. Define YOUR emergency change triggers and procedures",
        "4. Configure YOUR change risk classification matrix",
        "5. Document YOUR change calendar management approach",
        "6. Review the Summary_Dashboard for compliance metrics",
        "7. Maintain the Evidence Register for audit traceability",
        "8. Obtain final approval and sign-of",
    ]

    row += 1
    for instruction in instructions:
        ws[f"A{row}"] = instruction
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1

    # Status Legend
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "STATUS LEGEND"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    legend_headers = ["Symbol", "Status", "Description", "Color Code"]
    for col_idx, header in enumerate(legend_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    legend_data = [
        ("{CHECK}", "Defined", "Criteria/process clearly defined and documented", "Green"),
        ("{WARNING}", "Partial", "Partially defined or inconsistent application", "Yellow"),
        ("{XMARK}", "Not Defined", "Not defined or not documented", "Red"),
        ("📋", "Planned", "Definition planned with target date", "Blue"),
        ("N/A", "Not Applicable", "Not applicable to this environment", "Gray"),
    ]

    row += 1
    for symbol, status, desc, color in legend_data:
        ws[f"A{row}"] = symbol
        ws[f"B{row}"] = status
        ws[f"C{row}"] = desc
        ws[f"D{row}"] = color
        
        # Apply color coding
        if color == "Green":
            ws[f"D{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif color == "Yellow":
            ws[f"D{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif color == "Red":
            ws[f"D{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif color == "Blue":
            ws[f"D{row}"].fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        
        row += 1

    # Change Type Decision Tree
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "CHANGE TYPE DECISION TREE"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    decision_tree = [
        "Is this change:",
        "├─ Pre-approved, low-risk, well-documented? → STANDARD CHANGE",
        "├─ Requires assessment and approval? → NORMAL CHANGE",
        "│   ├─ Low Risk + Low Impact → Expedited Normal Change",
        "│   ├─ Medium Risk/Impact → Standard Normal Change",
        "│   └─ High/Critical Risk/Impact → High-Priority Normal Change",
        "└─ Urgent, meets emergency criteria? → EMERGENCY CHANGE",
        "    ├─ Critical system outage?",
        "    ├─ Security incident?",
        "    ├─ Data loss prevention?",
        "    └─ Regulatory deadline?",
    ]

    row += 1
    for line in decision_tree:
        ws[f"A{row}"] = line
        ws[f"A{row}"].font = Font(name="Courier New", size=9)
        row += 1

    # Acceptable Evidence
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (Examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    evidence_examples = [
        "✓ Standard changes catalog/library",
        "✓ Change classification decision trees",
        "✓ Risk assessment matrices",
        "✓ Emergency change criteria documentation",
        "✓ Change calendar procedures",
        "✓ CAB meeting records (classification decisions)",
        "✓ Change metrics (by type and category)",
        "✓ Training materials on change classification",
        "✓ Audit reports on classification accuracy",
        "✓ Exception/deviation records",
        "✓ Change type definitions document",
        "✓ Historical change data (by type)",
    ]

    row += 1
    for evidence in evidence_examples:
        ws[f"A{row}"] = evidence
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 4: STANDARD CHANGES CATALOG SHEET
# ============================================================================

def create_standard_changes_catalog(ws, styles):
    """Create Standard_Changes_Catalog sheet with 50 pre-formatted entries."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:M1")
    ws["A1"] = "STANDARD CHANGES CATALOG"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:M2")
    ws["A2"] = "Pre-approved, low-risk changes with documented procedures"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = ["Change ID", "Change Title", "Description", "Category", "Frequency", "Pre-requisites", 
               "Procedure Location", "Owner", "Approval Date", "Review Date", "Risk Level", "Status", "Evidence"]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Create 50 standard change rows
    row += 1
    for i in range(1, 51):
        # Change ID - suggested format
        ws[f"A{row}"] = f"STD-{i:03d}"
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Change Title, Description, Pre-requisites, Procedure Location, Owner, Evidence - text
        for col in ["B", "C", "F", "G", "H", "M"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Category dropdown
        validations['change_category'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Frequency dropdown
        validations['frequency'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Approval Date and Review Date
        for col in ["I", "J"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].number_format = 'DD.MM.YYYY'
        
        # Risk Level (Low/Medium only for standard changes)
        validations['std_change_risk'].add(ws[f"K{row}"])
        ws[f"K{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Status dropdown
        validations['std_change_status'].add(ws[f"L{row}"])
        ws[f"L{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Alternating row colors for readability
        if i % 2 == 0:
            for col in range(1, 14):
                cell = ws.cell(row=row, column=col)
                if cell.fill.start_color.rgb != "FFFFCC":
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        row += 1

    # Summary Metrics
    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "STANDARD CHANGE SUMMARY METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    metric_headers = ["Metric", "Value", "Notes"]
    for col_idx, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    metrics = [
        "Total Standard Changes Defined",
        "Active Standard Changes",
        "Standard Changes Under Review",
        "Standard Changes Requiring Annual Review",
        "Most Common Category",
        "Average Age of Standard Changes",
    ]

    row += 1
    for metric in metrics:
        ws[f"A{row}"] = metric
        ws[f"B{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        ws[f"B{row}"].value = "[FORMULA]"
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 18
    ws.column_dimensions["M"].width = 25

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 5: NORMAL CHANGES ASSESSMENT SHEET
# ============================================================================

def create_normal_changes_assessment(ws, styles):
    """Create Normal_Changes_Assessment sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "NORMAL CHANGES ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "Changes requiring risk assessment, approval, and CAB review"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== NORMAL CHANGE CRITERIA ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "NORMAL CHANGE CRITERIA"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    criteria_headers = ["Criterion", "Defined", "Documentation Reference", "Assessment Method", "Responsible Role", "Compliance", "Evidence"]
    for col_idx, header in enumerate(criteria_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    criteria = [
        "Change does not meet Standard criteria",
        "Change is not an emergency",
        "Risk assessment required",
        "Impact assessment required",
        "CAB review required (based on risk)",
        "Business justification required",
        "Implementation plan required",
        "Test plan required",
        "Rollback plan required",
        "Communication plan required",
        "PIR (Post-Implementation Review) required",
        "Change window scheduling required",
    ]

    row += 1
    for criterion in criteria:
        ws[f"A{row}"] = criterion
        
        # Editable cells
        for col in ["C", "D", "E", "G"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['definition_status'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['compliance_status'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== RISK-BASED APPROVAL PATHS ====================
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "RISK-BASED APPROVAL PATHS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    approval_headers = ["Risk Category", "Impact Category", "Approval Authority", "CAB Review", "Typical Timeline", "Success Rate", "Documented", "Evidence"]
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    risk_combinations = [
        ("Low", "Low"), ("Low", "Medium"), ("Low", "High"),
        ("Medium", "Low"), ("Medium", "Medium"), ("Medium", "High"),
        ("High", "Low"), ("High", "Medium"), ("High", "High"),
        ("Critical", "Any"),
    ]

    row += 1
    for risk, impact in risk_combinations:
        ws[f"A{row}"] = risk
        ws[f"B{row}"] = impact
        
        # Editable cells
        for col in ["E", "F", "H"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['approval_authority'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['cab_requirement'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['definition_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 25

    ws.freeze_panes = "A6"


# ============================================================================
# SECTION 6: EMERGENCY CHANGES SHEET
# ============================================================================

def create_emergency_changes(ws, styles):
    """Create Emergency_Changes sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "EMERGENCY CHANGES"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "Urgent changes meeting specific emergency criteria with E-CAB approval"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== EMERGENCY CRITERIA DEFINITION ====================
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "EMERGENCY CRITERIA DEFINITION"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    criteria_headers = ["Emergency Criterion", "Defined", "Specific Triggers", "Escalation Path", "Response Time SLA", "Documented", "Evidence"]
    # Adjust to 7 columns
    for col_idx, header in enumerate(criteria_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    emergency_criteria = [
        "System Outage (Critical Services)",
        "Security Incident Response",
        "Data Loss Prevention",
        "Regulatory Compliance Deadline",
        "Health & Safety Risk",
        "Business Continuity Threat",
        "[Custom Criterion 1]",
        "[Custom Criterion 2]",
    ]

    row += 1
    for criterion in emergency_criteria:
        ws[f"A{row}"] = criterion
        
        # Editable cells
        for col in ["C", "D", "E", "G"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['emergency_criteria_definition'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['definition_status'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== EMERGENCY CHANGE PROCESS REQUIREMENTS ====================
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "EMERGENCY CHANGE PROCESS REQUIREMENTS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    req_headers = ["Requirement", "Implemented", "Process Description", "Responsible Role", "Exceptions Allowed", "Compliance", "Evidence"]
    for col_idx, header in enumerate(req_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    requirements = [
        "Emergency criteria must be met",
        "E-CAB convened",
        "Minimum E-CAB members defined",
        "Emergency approval documented",
        "Risk acceptance explicit",
        "Communication immediate",
        "Implementation window immediate",
        "Rollback plan prepared (where feasible)",
        "Incident ticket linked",
        "Mandatory PIR within 2 business days",
        "Retrospective CAB review",
        "Emergency change metrics tracked",
    ]

    row += 1
    for requirement in requirements:
        ws[f"A{row}"] = requirement
        
        # Editable cells
        for col in ["C", "D", "G"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['implementation_level'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['exceptions_allowed'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['compliance_status'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== EMERGENCY CHANGE METRICS ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "EMERGENCY CHANGE METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    metric_headers = ["Metric", "Target", "Current (Last Quarter)", "Status", "Notes"]
    for col_idx, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    metrics = [
        "Emergency changes as % of total changes",
        "Average E-CAB response time",
        "Emergency changes with PIR completed",
        "Emergency changes leading to incidents",
        "Retrospective CAB review completion",
        "False emergency declarations",
        "Emergency change success rate",
    ]

    row += 1
    for metric in metrics:
        ws[f"A{row}"] = metric
        
        # Target and Notes - editable
        for col in ["B", "E"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Current - formula (editable for now)
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Status - calculated
        ws[f"D{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        ws[f"D{row}"].value = "[FORMULA]"
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 25

    ws.freeze_panes = "A6"


# ============================================================================
# SECTION 7: CHANGE RISK CLASSIFICATION SHEET
# ============================================================================

def create_change_risk_classification(ws, styles):
    """Create Change_Risk_Classification sheet with risk matrix."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "CHANGE RISK CLASSIFICATION MATRIX"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Risk = Impact × Likelihood methodology"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== IMPACT LEVEL DEFINITIONS ====================
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "IMPACT LEVEL DEFINITIONS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    impact_headers = ["Impact Level", "User Count Affected", "Service Downtime Potential", "Recovery Time", "Data Loss Risk", "Financial Impact", "Documented", "Evidence"]
    for col_idx, header in enumerate(impact_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    impact_levels = ["Low", "Medium", "High", "Critical"]

    row += 1
    for level in impact_levels:
        ws[f"A{row}"] = level
        
        # Editable cells
        for col in ["D", "F", "H"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['user_count'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['downtime_potential'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['data_loss_risk'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['definition_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== LIKELIHOOD LEVEL DEFINITIONS ====================
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "LIKELIHOOD LEVEL DEFINITIONS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    likelihood_headers = ["Likelihood Level", "Failure Probability", "Complexity", "Dependencies", "Testing Maturity", "Historical Success Rate", "Documented", "Evidence"]
    for col_idx, header in enumerate(likelihood_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    likelihood_levels = ["Low", "Medium", "High"]

    row += 1
    for level in likelihood_levels:
        ws[f"A{row}"] = level
        
        # Editable cells
        for col in ["B", "D", "F", "H"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['complexity'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['testing_maturity'].add(ws[f"E{row}"])
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['definition_status'].add(ws[f"G{row}"])
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== RISK MATRIX ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "RISK MATRIX (Impact × Likelihood)"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    # Risk matrix headers
    ws[f"A{row}"] = ""
    ws[f"B{row}"] = "Low Impact"
    ws[f"C{row}"] = "Medium Impact"
    ws[f"D{row}"] = "High Impact"
    ws[f"E{row}"] = "Critical Impact"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Risk matrix data with color coding
    risk_matrix = [
        ("Low Likelihood", "LOW RISK", "LOW RISK", "MEDIUM RISK", "HIGH RISK"),
        ("Medium Likelihood", "LOW RISK", "MEDIUM RISK", "HIGH RISK", "CRITICAL RISK"),
        ("High Likelihood", "MEDIUM RISK", "HIGH RISK", "CRITICAL RISK", "CRITICAL RISK"),
    ]

    row += 1
    for likelihood, low, medium, high, critical in risk_matrix:
        ws[f"A{row}"] = likelihood
        ws[f"A{row}"].font = Font(bold=True)
        
        ws[f"B{row}"] = low
        ws[f"C{row}"] = medium
        ws[f"D{row}"] = high
        ws[f"E{row}"] = critical
        
        # Apply risk color coding
        for col, value in [("B", low), ("C", medium), ("D", high), ("E", critical)]:
            if "LOW" in value:
                ws[f"{col}{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif "MEDIUM" in value:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif "HIGH" in value:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            elif "CRITICAL" in value:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"{col}{row}"].font = Font(bold=True)
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 25

    ws.freeze_panes = "A6"


# ============================================================================
# SECTION 8: CHANGE CALENDAR MANAGEMENT SHEET
# ============================================================================

def create_change_calendar_management(ws, styles):
    """Create Change_Calendar_Management sheet."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "CHANGE CALENDAR MANAGEMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "Scheduling, blackout windows, and conflict detection"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== CHANGE WINDOW DEFINITIONS ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "CHANGE WINDOW DEFINITIONS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    window_headers = ["Change Window", "Days/Times", "Applicable Change Types", "Approval Required", "Advance Notice", "Documented", "Evidence"]
    for col_idx, header in enumerate(window_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    change_windows = [
        "Standard Maintenance Window",
        "Business Hours (Restricted)",
        "After-Hours Window",
        "Emergency Window",
        "[Custom Window 1]",
        "[Custom Window 2]",
    ]

    row += 1
    for window in change_windows:
        ws[f"A{row}"] = window
        
        # Editable cells
        for col in ["B", "E", "G"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dropdowns
        validations['applicable_changes'].add(ws[f"C{row}"])
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['yes_no'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['definition_status'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== BLACKOUT PERIODS ====================
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "BLACKOUT PERIODS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    blackout_headers = ["Blackout Period", "Start Date", "End Date", "Reason", "Affected Systems/Services", "Exceptions Allowed", "Exception Approver", "Documented", "Evidence"]
    # Adjust to 9 columns
    for col_idx, header in enumerate(blackout_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    row += 1
    # 12 rows for blackout periods
    for i in range(12):
        # Blackout Period name - editable
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Start and End Date
        for col in ["B", "C"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].number_format = 'DD.MM.YYYY'
        
        # Affected Systems, Exception Approver, Evidence - editable
        for col in ["E", "G", "I"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Reason dropdown
        validations['blackout_reason'].add(ws[f"D{row}"])
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Exceptions Allowed dropdown
        validations['exceptions_allowed'].add(ws[f"F{row}"])
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Documented dropdown
        validations['definition_status'].add(ws[f"H{row}"])
        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 25

    ws.freeze_panes = "A6"


# ============================================================================
# SECTION 8.5: CLASSIFICATION METRICS SHEET
# ============================================================================

def create_classification_metrics(ws, styles):
    """Create Classification_Metrics sheet tracking change classification metrics."""
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "CHANGE CLASSIFICATION METRICS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Track metrics related to change type distribution and classification accuracy"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== CLASSIFICATION METRICS ====================
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "CHANGE TYPE DISTRIBUTION METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1

    headers = ["Metric", "Target", "Current", "Period", "Trend", "Status", "Owner", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    row += 1

    metrics = [
        ("Standard Changes %", ">60%", "", "Monthly", "", "", "", ""),
        ("Normal Changes %", "25-35%", "", "Monthly", "", "", "", ""),
        ("Emergency Changes %", "<5%", "", "Monthly", "", "", "", ""),
        ("Classification Accuracy", ">95%", "", "Monthly", "", "", "", ""),
        ("Re-Classification Rate", "<2%", "", "Monthly", "", "", "", ""),
        ("Standard Catalog Utilization", ">80%", "", "Monthly", "", "", "", ""),
    ]

    for metric_data in metrics:
        for col_idx, value in enumerate(metric_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 2:
                apply_style(cell, styles["input_cell"])
        row += 1

    row += 2

    # ==================== CLASSIFICATION ACCURACY ====================
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "CLASSIFICATION ACCURACY TRACKING"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1

    accuracy_headers = ["Period", "Total Changes", "Correctly Classified", "Misclassified", "Accuracy %", "Review Date", "Reviewer"]
    for col_idx, header in enumerate(accuracy_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    row += 1

    for _ in range(6):
        for col_idx in range(1, 8):
            cell = ws.cell(row=row, column=col_idx, value="")
            apply_style(cell, styles["input_cell"])
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 25

    ws.freeze_panes = "A6"


# ============================================================================
# SECTION 9: SUMMARY DASHBOARD SHEET
# ============================================================================

def create_summary_dashboard(ws, styles):
    """Create Summary_Dashboard with compliance metrics."""
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "CHANGE TYPES & CATEGORIES - SUMMARY DASHBOARD"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "Overall compliance status and key findings"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== OVERALL COMPLIANCE SUMMARY ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "OVERALL COMPLIANCE SUMMARY"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    summary_headers = ["Assessment Area", "Total Criteria", "Defined", "Partial", "Not Defined", "Compliance %", "Status"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    assessment_areas = [
        "Standard Changes Catalog",
        "Normal Changes Criteria",
        "Emergency Change Procedures",
        "Risk Classification Matrix",
        "Change Calendar Management",
    ]

    row += 1
    for area in assessment_areas:
        ws[f"A{row}"] = area
        
        # Formula cells (calculated)
        for col in range(2, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.value = "[FORMULA]"
            if col == 6:  # Compliance %
                cell.number_format = '0%'
        
        row += 1

    # Overall total row
    ws[f"A{row}"] = "OVERALL TOTAL"
    ws[f"A{row}"].font = Font(bold=True)
    for col in range(2, 8):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(bold=True)
        cell.value = "[FORMULA]"
        if col == 6:
            cell.number_format = '0%'

    # ==================== POLICY REQUIREMENT MAPPING ====================
    row += 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "POLICY REQUIREMENT MAPPING"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    req_headers = ["Policy Req ID", "Requirement Summary", "Status", "Evidence Sheet", "Evidence Row", "Auditor Notes"]
    for col_idx, header in enumerate(req_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    requirements = [
        ("REQ-CLASSIFY-001", "Standard changes catalog maintained"),
        ("REQ-CLASSIFY-002", "Standard change pre-approval"),
        ("REQ-CLASSIFY-003", "Normal change assessment criteria"),
        ("REQ-CLASSIFY-004", "Emergency criteria defined"),
        ("REQ-CLASSIFY-005", "Risk-based categorization"),
        ("REQ-CLASSIFY-006", "Impact assessment methodology"),
        ("REQ-CLASSIFY-007", "Likelihood assessment methodology"),
        ("REQ-CLASSIFY-008", "Risk matrix documented"),
        ("REQ-CLASSIFY-009", "Change calendar management"),
        ("REQ-CLASSIFY-010", "Blackout periods defined"),
        ("REQ-CLASSIFY-011", "E-CAB procedures"),
        ("REQ-CLASSIFY-012", "Classification metrics tracked"),
    ]

    row += 1
    for req_id, req_summary in requirements:
        ws[f"A{row}"] = req_id
        ws[f"B{row}"] = req_summary
        
        # Status - formula
        ws[f"C{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        ws[f"C{row}"].value = "[FORMULA]"
        
        # Evidence reference and notes - editable
        for col in ["D", "E", "F"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 15

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 10: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence_Register with 100 pre-formatted rows."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:I1")
    ws["A1"] = "EVIDENCE REGISTER"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = "Document all evidence supporting this assessment"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = ["Evidence ID", "Evidence Type", "Description", "Related Sheet/Requirement", "Location/Path", "Date Collected", "Collected By", "Verification Status", "Auditor Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # Create 100 evidence rows
    row += 1
    for i in range(1, 101):
        # Evidence ID - auto-generated
        ws[f"A{row}"] = f"EV-{i:03d}"
        ws[f"A{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Evidence Type - dropdown
        validations['evidence_type'].add(ws[f"B{row}"])
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Description, Related Sheet, Location, Collected By, Auditor Notes - text
        for col in ["C", "D", "E", "G", "I"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Date Collected - date
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].number_format = 'DD.MM.YYYY'
        
        # Verification Status - dropdown
        validations['verification_status'].add(ws[f"H{row}"])
        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Alternating row colors
        if i % 2 == 0:
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                if cell.fill.start_color.rgb not in ["FFFFCC", "E0E0E0"]:
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 30

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 11: APPROVAL SIGN-OFF SHEET
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create Approval_Sign_Off sheet with 3-level approval workflow."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "APPROVAL SIGN-OFF"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    # ==================== ASSESSMENT SUMMARY ====================
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    summary_fields = [
        ("Assessment Document:", "ISMS-IMP-A.8.32.2 - Change Types & Categories Assessment"),
        ("Assessment Period:", "[USER INPUT]"),
        ("Assessment Scope:", "[USER INPUT]"),
        ("Overall Compliance Rate:", "[FORMULA from Summary_Dashboard]"),
        ("Critical Gaps:", "[FORMULA from Summary_Dashboard]"),
        ("Audit Readiness:", "[FORMULA from Summary_Dashboard]"),
        ("Assessment Status:", "[DROPDOWN]"),
    ]

    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        
        if "USER INPUT" in value:
            ws[f"B{row}"] = ""
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        elif "DROPDOWN" in value:
            validations['assessment_status'].add(ws[f"B{row}"])
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        elif "FORMULA" in value:
            ws[f"B{row}"] = value
            ws[f"B{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        else:
            ws[f"B{row}"] = value
        
        row += 1

    # ==================== ASSESSMENT COMPLETED BY ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    preparer_fields = [
        "Name:", "Role/Title:", "Department:", "Email:", "Phone:", "Date Completed:", "Signature:"
    ]

    for field in preparer_fields:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        if field == "Date Completed:":
            ws[f"B{row}"].number_format = 'DD.MM.YYYY'
        
        row += 1

    # ==================== REVIEWED BY ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY - Change Manager"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    reviewer_fields = [
        ("Name:", None),
        ("Role/Title:", None),
        ("Review Date:", "date"),
        ("Review Notes:", "textarea"),
        ("Recommendation:", "dropdown"),
        ("Conditions (if any):", "text"),
        ("Signature:", None),
    ]

    for field, field_type in reviewer_fields:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        if field_type == "date":
            ws[f"B{row}"].number_format = 'DD.MM.YYYY'
        elif field_type == "dropdown":
            validations['review_recommendation'].add(ws[f"B{row}"])
        elif field_type == "textarea":
            ws.merge_cells(f"B{row}:E{row+2}")
            ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = 60
            row += 2
        
        row += 1

    # ==================== APPROVED BY ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY - CISO or IT Operations Manager"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    approver_fields = [
        ("Name:", None),
        ("Role/Title:", None),
        ("Approval Date:", "date"),
        ("Approval Decision:", "dropdown"),
        ("Conditions/Notes:", "textarea"),
        ("Signature:", None),
    ]

    for field, field_type in approver_fields:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        if field_type == "date":
            ws[f"B{row}"].number_format = 'DD.MM.YYYY'
        elif field_type == "dropdown":
            validations['approval_decision'].add(ws[f"B{row}"])
        elif field_type == "textarea":
            ws.merge_cells(f"B{row}:E{row+2}")
            ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = 60
            row += 2
        
        row += 1

    # ==================== NEXT REVIEW DETAILS ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    apply_style(ws[f"A{row}"], styles["section_header"])

    row += 1
    next_review_fields = [
        ("Next Review Date:", "formula"),
        ("Review Responsible:", None),
        ("Special Considerations:", "text"),
        ("Regulatory Review Required:", "dropdown"),
        ("External Audit Scheduled:", "date"),
    ]

    for field, field_type in next_review_fields:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        if field_type == "formula":
            ws[f"B{row}"] = "=TODAY()+90"
            ws[f"B{row}"].number_format = 'DD.MM.YYYY'
            ws[f"B{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        elif field_type == "date":
            ws[f"B{row}"].number_format = 'DD.MM.YYYY'
        elif field_type == "dropdown":
            validations['regulatory_review'].add(ws[f"B{row}"])
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Philosophy: Create evidence-based assessment tools for change classification,
    not checkbox compliance theater.
    """
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.32.2 - Change Types & Categories Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.32: Change Management")
    logger.info("=" * 78)
    logger.info("\n🎯 Systems Engineering Approach: Evidence-Based Compliance")
    logger.info(f"{CHART} Technology-Agnostic: Works with ANY change classification approach")
    logger.info(f"{LOCK} Audit-Ready: Comprehensive evidence collection")
    logger.info("\n" + "─" * 78)

    # Create workbook and setup styles
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    logger.info("{CHECK} Workbook created with 9 sheets")

    # Create all sheets
    logger.info("\n[Phase 2] Generating assessment sheets...")

    logger.info("  [1/10] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    logger.info("  ✅ Instructions complete")

    logger.info("  [2/10] Creating Standard_Changes_Catalog...")
    create_standard_changes_catalog(wb["Standard_Changes_Catalog"], styles)
    logger.info("  ✅ Standard changes catalog complete (50 entries)")

    logger.info("  [3/10] Creating Normal_Change_Classification...")
    create_normal_changes_assessment(wb["Normal_Change_Classification"], styles)
    logger.info("  ✅ Normal change classification complete")

    logger.info("  [4/10] Creating Emergency_Change_Procedures...")
    create_emergency_changes(wb["Emergency_Change_Procedures"], styles)
    logger.info("  ✅ Emergency change procedures complete")

    logger.info("  [5/10] Creating Risk_Assessment_Matrix...")
    create_change_risk_classification(wb["Risk_Assessment_Matrix"], styles)
    logger.info("  ✅ Risk assessment matrix complete")

    logger.info("  [6/10] Creating Change_Calendar_Management...")
    create_change_calendar_management(wb["Change_Calendar_Management"], styles)
    logger.info("  ✅ Change calendar and blackout periods complete")

    logger.info("  [7/10] Creating Classification_Metrics...")
    create_classification_metrics(wb["Classification_Metrics"], styles)
    logger.info("  ✅ Classification metrics complete")

    logger.info("  [8/10] Creating Evidence_Register...")
    create_evidence_register(wb["Evidence_Register"], styles)
    logger.info("  ✅ Evidence register complete (100 evidence rows)")

    logger.info("  [9/10] Creating Summary_Dashboard...")
    create_summary_dashboard(wb["Summary_Dashboard"], styles)
    logger.info("  ✅ Dashboard complete (compliance metrics)")

    logger.info("  [10/10] Creating Approval_Sign_Off...")
    create_approval_signoff(wb["Approval_Sign_Off"], styles)
    logger.info("  ✅ Approval workflow complete (3-level sign-off)")

    # Save workbook
    logger.info("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.32.2_Change_Types_Categories_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        wb.save(filename)
        logger.info("{CHECK} SUCCESS: {filename}")
    except Exception as e:
        logger.error("{XMARK} ERROR saving workbook: {e}")
        return 1

    # Summary
    logger.info("\n" + "=" * 78)
    logger.info("📋 WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 78)
    logger.info("\n📄 Assessment Sheets (10 per IMP specification):")
    logger.info("  • Instructions & Legend (usage guidance)")
    logger.info("  • Standard_Changes_Catalog (50 pre-approved changes)")
    logger.info("  • Normal_Change_Classification (risk-based approval paths)")
    logger.info("  • Emergency_Change_Procedures (E-CAB procedures, <5% target)")
    logger.info("  • Risk_Assessment_Matrix (Impact x Likelihood matrix)")
    logger.info("  • Change_Calendar_Management (blackout periods)")
    logger.info("  • Classification_Metrics (distribution and accuracy)")
    logger.info("\n📊 Analysis & Governance:")
    logger.info("  • Evidence_Register (100 evidence entries)")
    logger.info("  • Summary_Dashboard (compliance metrics)")
    logger.info("  • Approval_Sign_Off (3-level approval workflow)")
    logger.info("\n" + "─" * 78)
    logger.info("📈 ASSESSMENT CAPABILITIES:")
    logger.info("  • 50 standard changes catalog entries")
    logger.info("  • Risk matrix (4 impact × 3 likelihood levels)")
    logger.info("  • Emergency change metrics (<5% target)")
    logger.info("  • Change calendar with blackout periods")
    logger.info("  • 12 policy requirements mapped")
    logger.info("  • 100 evidence documentation entries")
    logger.info("  • Automated compliance calculations")
    logger.info("\n" + "─" * 78)
    logger.info(f"{TARGET} KEY FEATURES:")
    logger.info("  ✅ Technology-agnostic (works with ANY change management approach)")
    logger.info("  ✅ Standard/Normal/Emergency change types defined")
    logger.info("  ✅ Risk-based classification methodology")
    logger.info("  ✅ E-CAB procedures documented")
    logger.info("  ✅ Comprehensive evidence collection")
    logger.info("  ✅ Audit readiness assessment")
    logger.info("  ✅ Quarterly review cycle support")
    logger.info("\n" + "=" * 78)
    logger.info(f"{ROCKET} NEXT STEPS:")
    logger.info("  1. Open the generated workbook")
    logger.info("  2. Review Instructions & Legend sheet first")
    logger.info("  3. Populate Standard_Changes_Catalog with YOUR pre-approved changes")
    logger.info("  4. Document YOUR normal change criteria")
    logger.info("  5. Define YOUR emergency change triggers")
    logger.info("  6. Configure YOUR risk classification matrix")
    logger.info("  7. Document YOUR change calendar and blackout periods")
    logger.info("  8. Review Summary_Dashboard for compliance status")
    logger.info("  9. Document evidence in Evidence_Register")
    logger.info("  10. Obtain final approval via Approval_Sign_Off")
    logger.info("\n💡 PRO TIP:")
    logger.info("  If your emergency changes consistently stay below 5% of total changes")
    logger.info("  and your risk classifications accurately predict change outcomes,")
    logger.info("  you've moved beyond checkbox compliance to operational excellence.")
    logger.info("\n" + "=" * 78)
    logger.info('\n"The first principle is that you must not fool yourself')
    logger.info('— and you are the easiest person to fool." - Richard Feynman')
    logger.info("\n🎭 This is not cargo cult ISMS. This is evidence-based compliance.")
    logger.info("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
