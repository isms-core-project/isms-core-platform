#!/usr/bin/env python3
"""
normalize_assessment_files_a832.py

ISMS Assessment File Normalization Utility for Control A.8.32
Prepares assessment workbooks for dashboard linking

This script:
1. Scans directory for completed assessment files
2. Validates document IDs in Instructions & Legend sheet
3. Copies files to normalized filenames (no dates/versions)
4. Creates audit manifest for traceability

Usage:
    python3 normalize_assessment_files_a832.py

Requirements:
    sudo apt install python3-pip
    sudo apt install python3-openpyxl 
"""

import os
import sys
import shutil
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("ℹ️  Install with: sudo apt install python3-openpyxl")     
    sys.exit(1)


# ============================================================================
# CONFIGURATION
# ============================================================================

# Expected document IDs and titles from assessment workbooks
EXPECTED_DOCS = {
    "ISMS-IMP-A.8.32.1": {
        "title": "Change Process Assessment",
        "normalized": "ISMS-IMP-A.8.32.1.xlsx"
    },
    "ISMS-IMP-A.8.32.2": {
        "title": "Change Types & Categories Assessment",
        "normalized": "ISMS-IMP-A.8.32.2.xlsx"
    },
    "ISMS-IMP-A.8.32.3": {
        "title": "Environment Separation Assessment",
        "normalized": "ISMS-IMP-A.8.32.3.xlsx"
    },
    "ISMS-IMP-A.8.32.4": {
        "title": "Testing & Validation Assessment",
        "normalized": "ISMS-IMP-A.8.32.4.xlsx"
    },
}


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_workbook(filepath):
    """
    Validate workbook contains expected document ID in Instructions sheet.
    
    Args:
        filepath: Path to Excel workbook
    
    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        # Open workbook in read-only mode
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # Check for Instructions & Legend sheet (with or without underscore)
        sheet_name = None
        for name in ["Instructions & Legend", "Instructions_Legend", "Instructions"]:
            if name in wb.sheetnames:
                sheet_name = name
                break
        
        if not sheet_name:
            wb.close()
            return (None, None)
        
        ws = wb[sheet_name]
        
        # Look for Document ID in column A (typically rows 4-20)
        # Format: "Document ID" label in column A, value in column B
        for row in range(3, 25):
            cell_label = ws.cell(row=row, column=1).value
            
            if cell_label and "Document ID" in str(cell_label):
                doc_id_value = ws.cell(row=row, column=2).value
                
                if doc_id_value:
                    doc_id = str(doc_id_value).strip()
                    
                    # Check if it's one of our expected documents
                    if doc_id in EXPECTED_DOCS:
                        title = EXPECTED_DOCS[doc_id]["title"]
                        wb.close()
                        return (doc_id, title)
        
        wb.close()
        return (None, None)
        
    except Exception as e:
        print(f"    ⚠️  Error reading file: {e}")
        return (None, None)


def get_file_info(filepath):
    """Get file metadata for manifest."""
    stat = filepath.stat()
    return {
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime),
        "created": datetime.fromtimestamp(stat.st_ctime),
    }


# ============================================================================
# SCANNING FUNCTIONS
# ============================================================================

def scan_directory(directory):
    """
    Scan directory for assessment workbooks.
    
    Args:
        directory: Path to scan
    
    Returns:
        dict: {doc_id: {'path': Path, 'info': metadata}}
    """
    found_assessments = {}
    directory = Path(directory)
    
    # Get all .xlsx files, excluding temporary files
    xlsx_files = [f for f in directory.glob("*.xlsx") if not f.name.startswith("~$")]
    
    if not xlsx_files:
        print(f"\n❌ No Excel files found in {directory}\n")
        return found_assessments
    
    print(f"\n🔍 Scanning {len(xlsx_files)} Excel file(s) in {directory}...\n")
    
    for filepath in sorted(xlsx_files):
        print(f"  Checking: {filepath.name}")
        doc_id, title = validate_workbook(filepath)
        
        if doc_id:
            print(f"    ✅ Valid: {doc_id} - {title}")
            
            # Check for duplicates
            if doc_id in found_assessments:
                print(f"\n    ⚠️  WARNING: DUPLICATE FOUND FOR {doc_id}")
                print(f"        Previous: {found_assessments[doc_id]['path'].name}")
                print(f"        Current:  {filepath.name}\n")
                
                response = input("        Use CURRENT file? (y/n): ").strip().lower()
                
                if response == 'y':
                    found_assessments[doc_id] = {
                        'path': filepath,
                        'info': get_file_info(filepath)
                    }
                    print(f"        → Replaced with current file\n")
                else:
                    print(f"        → Keeping previous file\n")
            else:
                found_assessments[doc_id] = {
                    'path': filepath,
                    'info': get_file_info(filepath)
                }
        else:
            print(f"    ⏭️  Skipped (not a valid A.8.32 assessment workbook)")
        
        print()  # Blank line for readability
    
    return found_assessments


# ============================================================================
# MANIFEST CREATION
# ============================================================================

def create_manifest(output_dir, mapping, source_dir):
    """
    Create audit manifest documenting the normalization process.
    
    Args:
        output_dir: Destination directory
        mapping: Dictionary of normalized files
        source_dir: Source directory
    
    Returns:
        Path: Path to created manifest file
    """
    manifest_path = Path(output_dir) / "normalization_manifest.txt"
    
    with open(manifest_path, 'w', encoding='utf-8') as f:
        # Header
        f.write("=" * 80 + "\n")
        f.write("ISMS ASSESSMENT FILE NORMALIZATION MANIFEST\n")
        f.write("ISO/IEC 27001:2022 - Control A.8.32: Change Management\n")
        f.write("=" * 80 + "\n\n")
        
        # Normalization metadata
        f.write("NORMALIZATION METADATA\n")
        f.write("-" * 80 + "\n")
        f.write(f"Normalization Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Source Directory:        {source_dir.resolve()}\n")
        f.write(f"Output Directory:        {output_dir.resolve()}\n")
        f.write(f"Files Normalized:        {len(mapping)}/4 required\n")
        f.write(f"Normalization Status:    {'COMPLETE' if len(mapping) == 4 else 'INCOMPLETE'}\n")
        f.write("\n")
        
        # File mapping details
        f.write("FILE MAPPING\n")
        f.write("-" * 80 + "\n\n")
        
        for doc_id in sorted(EXPECTED_DOCS.keys()):
            if doc_id in mapping:
                info = mapping[doc_id]
                source_path = info['path']
                normalized_name = info['normalized']
                
                f.write(f"Document: {doc_id}\n")
                f.write(f"  Title:           {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"  Source File:     {source_path.name}\n")
                f.write(f"  Normalized Name: {normalized_name}\n")
                f.write(f"  File Size:       {info['info']['size']:,} bytes\n")
                f.write(f"  Modified:        {info['info']['modified'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"  Status:          ✅ NORMALIZED\n")
                f.write("\n")
            else:
                f.write(f"Document: {doc_id}\n")
                f.write(f"  Title:           {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"  Status:          ❌ NOT FOUND\n")
                f.write("\n")
        
        # Warnings
        if len(mapping) < 4:
            f.write("WARNINGS\n")
            f.write("-" * 80 + "\n")
            f.write(f"⚠️  Only {len(mapping)} of 4 required assessments were found and normalized.\n")
            f.write("   The compliance dashboard will have incomplete data for missing assessments.\n")
            f.write("\n   Missing assessments:\n")
            for doc_id in EXPECTED_DOCS.keys():
                if doc_id not in mapping:
                    f.write(f"   - {doc_id}: {EXPECTED_DOCS[doc_id]['title']}\n")
            f.write("\n")
        
        # Usage instructions
        f.write("NEXT STEPS\n")
        f.write("-" * 80 + "\n")
        f.write("1. Generate the compliance dashboard workbook:\n")
        f.write("   python3 generate_a832_5_compliance_dashboard.py\n\n")
        f.write("2. Place the dashboard workbook in this directory:\n")
        f.write(f"   {output_dir.resolve()}\n\n")
        f.write("3. Open the dashboard and click 'Update Links' when prompted\n\n")
        f.write("4. The dashboard will auto-populate with data from normalized files\n\n")
        
        f.write("HOW IT WORKS\n")
        f.write("-" * 80 + "\n")
        f.write("Dashboard Linking:\n")
        f.write("  - Dashboard contains formulas with external workbook references\n")
        f.write("  - External workbook links auto-update when sources change\n")
        f.write("  - Place dashboard in same directory as normalized files\n")
        f.write("  - Open dashboard and click 'Update Links' when prompted\n\n")
        f.write("File Retention:\n")
        f.write("  - Original source files: Retained in source directory\n")
        f.write("  - Normalized copies: Located in output directory\n")
        f.write("  - This manifest: Retained with normalized files for audit\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("END OF MANIFEST\n")
        f.write("=" * 80 + "\n")
    
    return manifest_path


# ============================================================================
# MAIN NORMALIZATION FUNCTION
# ============================================================================

def normalize_files(source_dir=None, output_dir=None, auto_confirm=False):
    """
    Main normalization orchestration function.
    
    Args:
        source_dir: Source directory (prompts if None)
        output_dir: Output directory (prompts if None)
        auto_confirm: Skip confirmation prompts (for automation)
    
    Returns:
        bool: True if successful, False otherwise
    """
    print("=" * 80)
    print("ISMS ASSESSMENT FILE NORMALIZATION UTILITY")
    print("ISO/IEC 27001:2022 - Control A.8.32: Change Management")
    print("=" * 80)
    print("\nThis script prepares assessment workbooks for dashboard linking by:")
    print("  1. Scanning for completed assessment files")
    print("  2. Validating document IDs in Instructions & Legend sheet")
    print("  3. Copying to normalized filenames (no dates/versions)")
    print("  4. Creating audit manifest for traceability\n")
    
    # Get source directory
    if not source_dir:
        source_dir = input("Enter source directory path (or press Enter for current directory): ").strip()
        if not source_dir:
            source_dir = "."
    
    source_dir = Path(source_dir).resolve()
    
    if not source_dir.exists():
        print(f"\n❌ Error: Source directory does not exist: {source_dir}\n")
        return False
    
    if not source_dir.is_dir():
        print(f"\n❌ Error: Not a directory: {source_dir}\n")
        return False
    
    print(f"\n🔍 Source directory: {source_dir}")
    
    # Get output directory
    if not output_dir:
        default_output = source_dir / "Dashboard_Sources"
        output_input = input(f"\nEnter output directory (or press Enter for '{default_output.name}'): ").strip()
        if not output_input:
            output_dir = default_output
        else:
            output_dir = Path(output_input)
    
    output_dir = Path(output_dir).resolve()
    
    # Create output directory if it doesn't exist
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        print(f"🔍 Output directory: {output_dir}")
    except Exception as e:
        print(f"\n❌ Error creating output directory: {e}\n")
        return False
    
    # Scan for assessment files
    found = scan_directory(source_dir)
    
    if not found:
        print("❌ No valid assessment workbooks found in source directory\n")
        print("   Ensure files contain valid Document IDs in 'Instructions & Legend' sheet:")
        for doc_id, info in EXPECTED_DOCS.items():
            print(f"   - {doc_id}: {info['title']}")
        print()
        return False
    
    # Show normalization summary
    print("=" * 80)
    print("NORMALIZATION SUMMARY")
    print("=" * 80 + "\n")
    print(f"Found {len(found)} of {len(EXPECTED_DOCS)} required assessment workbooks:\n")
    
    mapping = {}
    for doc_id in sorted(EXPECTED_DOCS.keys()):
        if doc_id in found:
            normalized_name = EXPECTED_DOCS[doc_id]["normalized"]
            source_path = found[doc_id]['path']
            
            print(f"  ✅ {doc_id}")
            print(f"     Title:      {EXPECTED_DOCS[doc_id]['title']}")
            print(f"     Source:     {source_path.name}")
            print(f"     Normalized: {normalized_name}\n")
            
            mapping[doc_id] = {
                'path': source_path,
                'normalized': normalized_name,
                'info': found[doc_id]['info']
            }
        else:
            print(f"  ❌ {doc_id} - NOT FOUND")
            print(f"     Title: {EXPECTED_DOCS[doc_id]['title']}\n")
    
    if len(found) < len(EXPECTED_DOCS):
        print(f"⚠️  WARNING: Only {len(found)}/{len(EXPECTED_DOCS)} assessment workbooks found")
        print("   Dashboard will have incomplete data for missing assessments\n")
    
    print(f"Output directory: {output_dir}\n")
    
    # Confirm normalization
    if not auto_confirm:
        response = input("Proceed with normalization? (y/n): ").strip().lower()
        if response != 'y':
            print("\n❌ Normalization cancelled by user\n")
            return False
    
    # Perform normalization (copy files)
    print("\n" + "=" * 80)
    print("NORMALIZING FILES...")
    print("=" * 80 + "\n")
    
    for doc_id, info in mapping.items():
        source = info['path']
        dest = output_dir / info['normalized']
        
        print(f"Copying: {source.name}")
        print(f"      → {dest.name}")
        
        try:
            shutil.copy2(source, dest)
            print(f"      ✅ Success\n")
        except Exception as e:
            print(f"      ❌ Error: {e}\n")
            print(f"❌ Normalization failed at {doc_id}\n")
            return False
    
    # Create audit manifest
    print("📄 Creating audit manifest...")
    try:
        manifest = create_manifest(output_dir, mapping, source_dir)
        print(f"   ✅ Created: {manifest.name}\n")
    except Exception as e:
        print(f"   ❌ Error creating manifest: {e}\n")
        return False
    
    # Success summary
    print("=" * 80)
    print("✅ NORMALIZATION COMPLETE")
    print("=" * 80 + "\n")
    print(f"Normalized files:  {output_dir}")
    print(f"Audit manifest:    {manifest}\n")
    print("NEXT STEPS:\n")
    print("  1. Review audit manifest for file mapping details")
    print("  2. Generate dashboard workbook:")
    print("     python3 generate_a832_5_compliance_dashboard.py\n")
    print("  3. Place dashboard in same directory as normalized files:")
    print(f"     {output_dir}\n")
    print("  4. Open dashboard and click 'Update Links' when prompted\n")
    print("  5. Dashboard will auto-populate with current compliance data\n")
    print("=" * 80 + "\n")
    
    return True


# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Normalize ISMS assessment workbooks for dashboard integration",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode (prompts for directories)
  python3 normalize_assessment_files_a832.py
  
  # Specify source directory
  python3 normalize_assessment_files_a832.py --source ./assessments
  
  # Specify both directories
  python3 normalize_assessment_files_a832.py --source ./assessments --output ./dashboard
  
  # Automated mode (no prompts)
  python3 normalize_assessment_files_a832.py --source ./assessments --auto-confirm
        """
    )
    
    parser.add_argument(
        '--source', '-s',
        help='Source directory containing assessment workbooks (default: current directory)'
    )
    
    parser.add_argument(
        '--output', '-o',
        help='Output directory for normalized files (default: ./Dashboard_Sources)'
    )
    
    parser.add_argument(
        '--auto-confirm', '-y',
        action='store_true',
        help='Skip confirmation prompts (for automation)'
    )
    
    args = parser.parse_args()
    
    success = normalize_files(
        source_dir=args.source,
        output_dir=args.output,
        auto_confirm=args.auto_confirm
    )
    
    sys.exit(0 if success else 1)