#!/usr/bin/env python3
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
Excel Workbook Sanity Checker for ISMS Control A.8.9 Configuration Management
Diagnoses issues that trigger Excel's "file level validation and repair"

GENERIC VERSION - Works for all A.8.9 assessment workbooks (A.8.9.1 through A.8.9.5)

Usage:
    python3 excel_sanity_check_a89.py ISMS_A_8_9_1_Baseline_Configuration_Assessment_20260108.xlsx
    python3 excel_sanity_check_a89.py ISMS_A_8_9_5_Compliance_Dashboard_20260108.xlsx

"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)



# =============================================================================
# WORKBOOK TYPE DETECTION
# =============================================================================

def detect_workbook_type(filename):
    """Detect which A.8.9 workbook type this is based on filename."""
    
    if 'A_8_9_1' in filename or 'A.8.9.1' in filename:
        return 'A.8.9.1', 'Baseline Configuration Assessment'
    elif 'A_8_9_2' in filename or 'A.8.9.2' in filename:
        return 'A.8.9.2', 'Change Control Assessment'
    elif 'A_8_9_3' in filename or 'A.8.9.3' in filename:
        return 'A.8.9.3', 'Configuration Monitoring Assessment'
    elif 'A_8_9_4' in filename or 'A.8.9.4' in filename:
        return 'A.8.9.4', 'Security Hardening Assessment'
    elif 'A_8_9_5' in filename or 'A.8.9.5' in filename:
        return 'A.8.9.5', 'Compliance Dashboard'
    else:
        return 'Unknown', 'Generic Excel Workbook'


# =============================================================================
# HEALTH CHECK FUNCTIONS
# =============================================================================

def check_workbook_health(filename):
    """
    Perform comprehensive health check on Excel workbook.
    
    Args:
        filename: Path to Excel file
    """
    print(f"\n{'=' * 70}")
    print(f"Excel Sanity Check - ISMS Control A.8.9 Configuration Management")
    print(f"{'=' * 70}\n")
    print(f"File: {filename}\n")
    
    workbook_type, workbook_title = detect_workbook_type(filename)
    print(f"Detected Type: {workbook_type} - {workbook_title}\n")
    
    # Load workbook
    try:
        print("Loading workbook (read-only mode)...")
        wb = load_workbook(filename, read_only=False, data_only=False)
        print("✅ Workbook loaded successfully\n")
    except Exception as e:
        print(f"✗ ERROR loading workbook: {e}\n")
        return
    
    # Basic workbook info
    print(f"Sheet Count: {len(wb.sheetnames)}")
    print(f"Sheets: {', '.join(wb.sheetnames)}\n")
    
    issues_found = 0
    warnings_found = 0
    
    # Check 1: Sheet names
    print("=" * 70)
    print("CHECK 1: Sheet Name Validation")
    print("=" * 70)
    
    for sheet_name in wb.sheetnames:
        if len(sheet_name) > 31:
            print(f"✗ ERROR: Sheet name too long (>31 chars): '{sheet_name}'")
            issues_found += 1
        
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for char in invalid_chars:
            if char in sheet_name:
                print(f"✗ ERROR: Invalid character '{char}' in sheet name: '{sheet_name}'")
                issues_found += 1
    
    if issues_found == 0:
        print("✅ All sheet names valid")
    print()
    
    # Check 2: Formula validation
    print("=" * 70)
    print("CHECK 2: Formula Validation")
    print("=" * 70)
    
    formula_issues = 0
    external_refs = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula cell
                    formula = str(cell.value)
                    
                    # Check for external references
                    if '[' in formula and ']' in formula:
                        external_refs += 1
                    
                    # Check for common formula errors
                    if formula.count('(') != formula.count(')'):
                        print(f"⚠  WARNING: Unbalanced parentheses in {sheet_name}!{cell.coordinate}: {formula[:50]}")
                        warnings_found += 1
                        formula_issues += 1
                    
                    if '#REF!' in formula:
                        print(f"✗ ERROR: #REF! error in {sheet_name}!{cell.coordinate}")
                        issues_found += 1
                        formula_issues += 1
    
    if formula_issues == 0:
        print(f"✓ No formula issues detected")
    
    if external_refs > 0:
        print(f"ℹ  Found {external_refs} external workbook references")
        if workbook_type == 'A.8.9.5':
            print(f"   This is EXPECTED for the Compliance Dashboard")
        else:
            print(f"⚠  WARNING: External references in non-dashboard workbook")
            warnings_found += 1
    print()
    
    # Check 3: Data validation
    print("=" * 70)
    print("CHECK 3: Data Validation Rules")
    print("=" * 70)
    
    validation_count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, 'data_validations') and ws.data_validations:
            validation_count += len(ws.data_validations.dataValidation)
    
    print(f"✓ Found {validation_count} data validation rules")
    print()
    
    # Check 4: Conditional formatting
    print("=" * 70)
    print("CHECK 4: Conditional Formatting")
    print("=" * 70)
    
    cf_count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, 'conditional_formatting'):
            cf_count += len(ws.conditional_formatting._cf_rules)
    
    print(f"✓ Found {cf_count} conditional formatting rules")
    print()
    
    # Check 5: Cell protection
    print("=" * 70)
    print("CHECK 5: Sheet Protection")
    print("=" * 70)
    
    protected_sheets = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.protection.sheet:
            protected_sheets += 1
    
    print(f"✓ {protected_sheets}/{len(wb.sheetnames)} sheets are protected")
    if protected_sheets == 0 and workbook_type != 'Unknown':
        print(f"⚠  WARNING: No sheets protected - expected for ISMS workbooks")
        warnings_found += 1
    print()
    
    # Check 6: Large data sets
    print("=" * 70)
    print("CHECK 6: Data Volume Analysis")
    print("=" * 70)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row > 10000:
            print(f"⚠  WARNING: {sheet_name} has {max_row} rows (may be slow)")
            warnings_found += 1
        
        if max_col > 50:
            print(f"⚠  WARNING: {sheet_name} has {max_col} columns (may be slow)")
            warnings_found += 1
    
    print(f"✓ Data volume check complete")
    print()
    
    # Check 7: Workbook-specific validation
    print("=" * 70)
    print("CHECK 7: Workbook-Specific Validation")
    print("=" * 70)
    
    if workbook_type == 'A.8.9.5':
        # Dashboard-specific checks
        print("Compliance Dashboard Checks:")
        
        required_sheets = [
            'Instructions',
            'Workbook_Integration_Settings',
            'Overall_Compliance_Dashboard'
        ]
        
        missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]
        if missing_sheets:
            print(f"✗ ERROR: Missing required sheets: {', '.join(missing_sheets)}")
            issues_found += len(missing_sheets)
        else:
            print(f"✓ All required dashboard sheets present")
        
        # Check for external references
        if external_refs == 0:
            print(f"✗ ERROR: Dashboard has NO external references")
            print(f"   Dashboard REQUIRES links to 4 source workbooks")
            issues_found += 1
        else:
            print(f"✓ External references found ({external_refs})")
    
    elif workbook_type != 'Unknown':
        # Assessment workbook checks
        print(f"{workbook_title} Checks:")
        
        if 'Instructions' not in wb.sheetnames:
            print(f"⚠  WARNING: No Instructions sheet found")
            warnings_found += 1
        else:
            print(f"✓ Instructions sheet present")
        
        if 'Approval_Sign_Off' not in wb.sheetnames and 'Approval Sign-Off' not in wb.sheetnames:
            print(f"⚠  WARNING: No Approval sign-off sheet found")
            warnings_found += 1
        else:
            print(f"✓ Approval sign-off sheet present")
    
    print()
    
    # Final summary
    print("=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"Workbook Type: {workbook_type} - {workbook_title}")
    print(f"Total Sheets: {len(wb.sheetnames)}")
    print(f"Formulas with External References: {external_refs}")
    print(f"Data Validation Rules: {validation_count}")
    print(f"Conditional Formatting Rules: {cf_count}")
    print(f"Protected Sheets: {protected_sheets}")
    print()
    print(f"Issues Found: {issues_found}")
    print(f"Warnings: {warnings_found}")
    print()
    
    if issues_found == 0 and warnings_found == 0:
        print("✅✅✅ WORKBOOK HEALTH: EXCELLENT")
        print("    No issues detected. File should open without errors.")
    elif issues_found == 0:
        print("✅ WORKBOOK HEALTH: GOOD")
        print("  No critical issues. Warnings are informational.")
    else:
        print("❌ WORKBOOK HEALTH: ISSUES DETECTED")
        print(f"  {issues_found} issue(s) may cause Excel to prompt for repair.")
    
    print()
    
    # Workbook-specific guidance
    if workbook_type == 'A.8.9.5':
        print("DASHBOARD-SPECIFIC GUIDANCE:")
        print("  • Dashboard REQUIRES 4 source workbooks in same directory")
        print("  • Source workbooks must have normalized names (no dates)")
        print("  • Run normalize_assessment_files_a89.py first")
        print("  • Open dashboard and click 'Update Links' when prompted")
        print("  • If #REF! errors: Check Workbook_Integration_Settings sheet")
    
    print("\n" + "=" * 70 + "\n")
    
    wb.close()


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("=" * 70)
        print("ISMS Control A.8.9 Configuration Management - Excel Sanity Checker")
        print("=" * 70)
        print("\nUsage: python3 excel_sanity_check_a89.py <filename.xlsx>")
        print("\nExamples:")
        print("  python3 excel_sanity_check_a89.py ISMS_A_8_9_1_Baseline_Configuration_Assessment_20260108.xlsx")
        print("  python3 excel_sanity_check_a89.py ISMS_A_8_9_5_Compliance_Dashboard_20260108.xlsx")
        print("\nSupported workbook types:")
        print("  A.8.9.1 - Baseline Configuration Assessment")
        print("  A.8.9.2 - Change Control Assessment")
        print("  A.8.9.3 - Configuration Monitoring Assessment")
        print("  A.8.9.4 - Security Hardening Assessment")
        print("  A.8.9.5 - Compliance Dashboard")
        print("\n'Evidence > Theater' - Systems Engineering ISMS")
        sys.exit(1)
    
    filename = sys.argv[1]
    
    try:
        check_workbook_health(filename)
    except FileNotFoundError:
        print(f"\n✗ ERROR: File not found: {filename}")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (syntax validated, structure verified)
# QA_TOOL: Claude Code Deep Scan
# QA_NOTE: STANDARDIZATION - License header, logging, main() pattern applied
# =============================================================================
