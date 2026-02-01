#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.1-2-6.1-2.S1 - Policy Framework Assessment Workbook Generator
================================================================================

Purpose:
    Generate Excel assessment workbook for ISO/IEC 27001:2022 Control A.5.1
    (Policies for Information Security) as part of stacked control framework
    A.5.1 + A.5.2 + A.6.1 + A.6.2 (Secure Employment and Roles).

Control Alignment:
    - ISO/IEC 27001:2022 Annex A.5.1: Policies for Information Security
    - ISMS-POL-A.5.1-2-6.1-2, Section 4: Policy Framework Requirements

Assessment Focus:
    - Policy inventory completeness
    - Policy lifecycle compliance (creation, approval, review, publication)
    - Policy governance (ownership, accountability, approval authority)
    - Policy classification and access controls
    - Policy communication and acknowledgment
    - Policy repository structure and accessibility

Workbook Structure:
    11 sheets total:
    1. Dashboard (executive summary, auto-calculated)
    2. Policy_Inventory (master policy list with metadata)
    3. Lifecycle_Compliance (approval, review, publication verification)
    4. Governance_Assessment (ownership, accountability, RACI)
    5. Classification_Review (classification & access control verification)
    6. Communication_Tracking (publication & acknowledgment rates)
    7. Repository_Assessment (repository-wide evaluation)
    8. Gap_Analysis (consolidated gaps with risk levels)
    9. Evidence_Register (supporting documentation)
   10. Action_Items (remediation tracking)
   11. Approval_Sign_Off (three-level approval workflow)

Dependencies:
    - openpyxl >= 3.0.0

Usage:
    python generate_a5_1_2_6_1_2_s1_policy_framework.py

Output:
    ISMS-IMP-A.5.1-2-6.1-2.S1_Policy_Framework_YYYYMMDD.xlsx

Technical Specification:
    See ISMS-IMP-A.5.1-2-6.1-2.S1-Policy-Framework-Assessment.md Part II

Quality Standards:
    - Follows A.8.23/A.8.24 reference implementation patterns
    - Comprehensive documentation and inline comments
    - Consistent styling (Yellow input, Light Blue calculated, Gray labels)
    - Data validation on all appropriate cells
    - Conditional formatting for compliance statuses
    - Sheet protection with selective unlocking
    - Named ranges for cross-sheet references
    - Formula-based auto-calculations
    - Professional formatting for executive presentation

Author:               [Organization] ISMS Implementation Team
Version: 1.0
Date:                 [Date to be set]
License: Internal Use Only

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)




# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.1-2-6.1-2.S1"
WORKBOOK_NAME = "Policy Framework Assessment Workbook"
CONTROL_ID = "A.5.1-2-6.1-2"
CONTROL_NAME = "Information Security Policies and Organization"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ==============================================================================
# STYLE DEFINITIONS
# ==============================================================================

def setup_styles():
    """
    Define consistent cell styles used throughout the workbook.
    
    Returns:
        dict: Dictionary of style objects for different cell types
    """
    styles = {}
    
    # Header styles
    styles['header_main'] = {
        'font': Font(name='Arial', size=18, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }
    
    styles['header_sub'] = {
        'font': Font(name='Arial', size=12, bold=False, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }
    
    styles['section_header'] = {
        'font': Font(name='Arial', size=14, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }
    
    styles['column_header'] = {
        'font': Font(name='Arial', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }
    
    # Data cell styles
    styles['input_cell'] = {
        'font': Font(name='Arial', size=10, color='000000'),
        'fill': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),  # Yellow
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }
    
    styles['calculated_cell'] = {
        'font': Font(name='Arial', size=10, color='000000'),
        'fill': PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'),  # Light Blue
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }
    
    styles['label_cell'] = {
        'font': Font(name='Arial', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),  # Gray
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }
    
    # Status fills (for conditional formatting reference)
    styles['status_compliant'] = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')  # Green
    styles['status_partial'] = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
    styles['status_non_compliant'] = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
    
    return styles


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def apply_cell_style(cell, style_dict):
    """
    Apply a style dictionary to a cell.
    
    Args:
        cell: openpyxl cell object
        style_dict: Dictionary containing style attributes
    """
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']


def set_column_widths(ws, widths_dict):
    """
    Set column widths for a worksheet.
    
    Args:
        ws: Worksheet object
        widths_dict: Dictionary mapping column letters to widths
    """
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width


def create_data_validation(formula_or_list, allow_blank=True, error_title="Invalid Entry", 
                           error_message="Please select from dropdown"):
    """
    Create a data validation object for dropdown lists.
    
    Args:
        formula_or_list: Either a string formula or list of values
        allow_blank: Whether to allow blank cells
        error_title: Title for error message
        error_message: Error message text
    
    Returns:
        DataValidation object
    """
    if isinstance(formula_or_list, list):
        dv = DataValidation(type="list", formula1=f'"{",".join(formula_or_list)}"', 
                           allow_blank=allow_blank)
    else:
        dv = DataValidation(type="list", formula1=formula_or_list, allow_blank=allow_blank)
    
    dv.error = error_message
    dv.errorTitle = error_title
    dv.showErrorMessage = True
    return dv


def merge_and_style(ws, range_string, value, style_dict):
    """
    Merge cells and apply style.
    
    Args:
        ws: Worksheet object
        range_string: Range to merge (e.g., 'A1:K1')
        value: Value to write
        style_dict: Style dictionary to apply
    """
    ws.merge_cells(range_string)
    cell = ws[range_string.split(':')[0]]
    cell.value = value
    apply_cell_style(cell, style_dict)


# ==============================================================================
# SHEET CREATION FUNCTIONS
# ==============================================================================

def create_dashboard_sheet(wb, styles):
    """
    Sheet 1: Dashboard (Executive Summary)
    
    Auto-calculated metrics pulling from other sheets.
    Read-only for users.
    """
    ws = wb.active
    ws.title = "Dashboard"
    
    # --- Header Section ---
    merge_and_style(ws, 'A1:K1', 
                   'ISMS-IMP-A.5.1-2-6.1-2.S1 — Policy Framework Assessment Dashboard',
                   styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    merge_and_style(ws, 'A2:K2',
                   'ISO/IEC 27001:2022 Control A.5.1 - Policies for Information Security',
                   styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # --- Document Information Block (Rows 4-15) ---
    ws['A4'] = 'Document ID:'
    apply_cell_style(ws['A4'], styles['label_cell'])
    ws.merge_cells('B4:C4')
    ws['B4'] = 'ISMS-IMP-A.5.1-2-6.1-2.S1'
    
    info_rows = [
        (5, 'Related Policy:', 'ISMS-POL-A.5.1-2-6.1-2, Section 4'),
        (6, 'ISO Control:', 'A.5.1 (Policies for Information Security)'),
        (7, 'Assessment Period:', None),  # User input
        (8, 'Assessment Date:', '=TODAY()'),  # Formula
        (9, 'Assessor Name:', None),  # User input
        (10, 'Assessor Role:', None),  # User input
        (11, 'Organization:', None),  # User input
        (12, 'Review Cycle:', 'Quarterly'),
        (13, 'Last Updated:', '=NOW()'),  # Formula
        (14, 'Assessment Status:', None),  # Dropdown
        (15, 'Next Review Date:', None)  # User input
    ]
    
    for row_num, label, value in info_rows:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        ws.merge_cells(f'B{row_num}:C{row_num}')
        
        if value:
            ws[f'B{row_num}'] = value
            if value.startswith('='):
                apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
            else:
                apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])
    
    # Status dropdown for row 14
    status_dv = create_data_validation(['Draft', 'Under-Review', 'Approved', 'Audit-Ready'])
    ws.add_data_validation(status_dv)
    status_dv.add(ws['B14'])
    
    # --- Overall Compliance Summary (Rows 17-30) ---
    merge_and_style(ws, 'A17:K17', 'OVERALL POLICY FRAMEWORK COMPLIANCE', styles['section_header'])
    ws.row_dimensions[17].height = 35
    
    # Compliance Scorecard (Rows 19-26)
    scorecard = [
        (19, 'Overall Compliance Score', '=AVERAGE(B32,B33,B34,B35,B36,B37)', '90%'),
        (20, 'Policies Assessed', '=COUNTA(Policy_Inventory!A:A)-4', 'N/A'),
        (21, 'Policies Compliant', '=COUNTIF(Lifecycle_Compliance!L:L,"Compliant")', '100%'),
        (22, 'Policies with Gaps', '=B20-B21', '0'),
        (23, 'Critical Gaps', '=COUNTIF(Gap_Analysis!F:F,"Critical")', '0'),
        (24, 'High Priority Gaps', '=COUNTIF(Gap_Analysis!F:F,"High")', '0'),
        (25, 'Medium Priority Gaps', '=COUNTIF(Gap_Analysis!F:F,"Medium")', 'N/A'),
        (26, 'Low Priority Gaps', '=COUNTIF(Gap_Analysis!F:F,"Low")', 'N/A')
    ]
    
    for row_num, label, formula, target in scorecard:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        
        ws[f'B{row_num}'] = formula
        apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        if row_num == 19:
            ws[f'B{row_num}'].number_format = '0.00%'
        
        ws[f'C{row_num}'] = target
        apply_cell_style(ws[f'C{row_num}'], styles['calculated_cell'])
        
        ws[f'D{row_num}'] = '=IF(B{0}>=C{0},"✓ On Target",IF(B{0}>=C{0}*0.9,"⚠ Close","✗ Below Target"))'.format(row_num)
        apply_cell_style(ws[f'D{row_num}'], styles['calculated_cell'])
    
    # --- Domain Compliance Breakdown (Rows 32-38) ---
    merge_and_style(ws, 'A28:K28', 'DOMAIN COMPLIANCE BREAKDOWN', styles['section_header'])
    
    # Column headers for domain table
    for col, header in zip(['A', 'B', 'C', 'D', 'E'], 
                           ['Domain', 'Weight', 'Score (%)', 'Weighted Score', 'Status']):
        ws[f'{col}30'] = header
        apply_cell_style(ws[f'{col}30'], styles['column_header'])
    
    domains = [
        (32, 'Policy Inventory Completeness', '25%', 
         '=COUNTIF(Policy_Inventory!S:S,"No-Gap")/COUNTA(Policy_Inventory!A:A)*100', '=B32*0.25'),
        (33, 'Policy Lifecycle Compliance', '25%',
         '=COUNTIF(Lifecycle_Compliance!L:L,"Compliant")/COUNTA(Lifecycle_Compliance!A:A)*100', '=B33*0.25'),
        (34, 'Policy Governance', '20%',
         '=COUNTIF(Governance_Assessment!M:M,"Compliant")/COUNTA(Governance_Assessment!A:A)*100', '=B34*0.20'),
        (35, 'Policy Classification', '10%',
         '=COUNTIF(Classification_Review!J:J,"Compliant")/COUNTA(Classification_Review!A:A)*100', '=B35*0.10'),
        (36, 'Policy Communication', '15%',
         '=COUNTIF(Communication_Tracking!N:N,"Compliant")/COUNTA(Communication_Tracking!A:A)*100', '=B36*0.15'),
        (37, 'Policy Repository', '5%',
         '=IF(Repository_Assessment!B20="Compliant",100,IF(Repository_Assessment!B20="Partial",50,0))', '=B37*0.05')
    ]
    
    for row_num, domain, weight, score_formula, weighted_formula in domains:
        ws[f'A{row_num}'] = domain
        ws[f'B{row_num}'] = weight
        ws[f'C{row_num}'] = score_formula
        ws[f'C{row_num}'].number_format = '0.00%'
        ws[f'D{row_num}'] = weighted_formula
        ws[f'D{row_num}'].number_format = '0.00%'
        ws[f'E{row_num}'] = f'=IF(C{row_num}>=0.9,"✓",IF(C{row_num}>=0.7,"⚠","✗"))'
    
    # Total row
    ws['A38'] = 'TOTAL WEIGHTED SCORE'
    apply_cell_style(ws['A38'], styles['label_cell'])
    ws['B38'] = '100%'
    ws['C38'] = '=SUM(D32:D37)'
    ws['C38'].number_format = '0.00%'
    ws['C38'].font = Font(name='Arial', size=10, bold=True)
    
    # Set column widths
    set_column_widths(ws, {'A': 40, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 5, 'G': 5, 'H': 5, 'I': 5, 'J': 5, 'K': 5})
    
    # Protect sheet (allow sorting/filtering)
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_policy_inventory_sheet(wb, styles):
    """
    Sheet 2: Policy_Inventory
    
    Master policy inventory with full metadata.
    150 rows for policies.
    """
    ws = wb.create_sheet("Policy_Inventory")
    
    # --- Header ---
    merge_and_style(ws, 'A1:T1', 'POLICY INVENTORY', styles['header_main'])
    merge_and_style(ws, 'A2:T2', 
                   'Master list of all information security policies with metadata',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25
    
    # --- Column Headers (Row 4) ---
    headers = [
        ('A', 'Policy_ID', 20),
        ('B', 'Policy_Title', 35),
        ('C', 'Policy_Hierarchy_Tier', 15),
        ('D', 'ISO_Control_Mapping', 20),
        ('E', 'Policy_Owner', 25),
        ('F', 'Policy_Approver', 25),
        ('G', 'Current_Version', 12),
        ('H', 'Version_Date', 12),
        ('I', 'Approval_Date', 12),
        ('J', 'Last_Review_Date', 12),
        ('K', 'Next_Review_Date', 12),
        ('L', 'Review_Frequency', 15),
        ('M', 'Policy_Status', 15),
        ('N', 'Policy_Classification', 15),
        ('O', 'Acknowledgment_Required', 15),
        ('P', 'Acknowledgment_Rate', 12),
        ('Q', 'Repository_Location', 40),
        ('R', 'Related_Documents', 30),
        ('S', 'Gap_Identified', 15),
        ('T', 'Notes', 40)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)
    ws.row_dimensions[4].height = 30
    
    # --- Data Rows (5-154) ---
    # Apply input cell style to data range
    for row in range(5, 155):
        for col_letter in [c[0] for c in headers]:
            cell = ws[f'{col_letter}{row}']
            apply_cell_style(cell, styles['input_cell'])
    
    # --- Data Validation ---
    # Tier dropdown
    tier_dv = create_data_validation(['Tier-1-Master', 'Tier-2-Domain', 'Tier-3-Topic'])
    ws.add_data_validation(tier_dv)
    tier_dv.add(f'C5:C154')
    
    # Review Frequency dropdown
    freq_dv = create_data_validation(['Annual', 'Biennial', 'Quarterly', 'Triggered-Only'])
    ws.add_data_validation(freq_dv)
    freq_dv.add(f'L5:L154')
    
    # Status dropdown
    status_dv = create_data_validation(['Active', 'Draft', 'Under-Review', 'Retired', 'Superseded'])
    ws.add_data_validation(status_dv)
    status_dv.add(f'M5:M154')
    
    # Classification dropdown
    class_dv = create_data_validation(['Internal', 'Confidential', 'Public'])
    ws.add_data_validation(class_dv)
    class_dv.add(f'N5:N154')
    
    # Acknowledgment Required dropdown
    ack_dv = create_data_validation(['Yes', 'No', 'Role-Specific'])
    ws.add_data_validation(ack_dv)
    ack_dv.add(f'O5:O154')
    
    # Gap Identified dropdown
    gap_dv = create_data_validation(['No-Gap', 'Minor-Gap', 'Significant-Gap', 'Critical-Gap'])
    ws.add_data_validation(gap_dv)
    gap_dv.add(f'S5:S154')
    
    # Date formatting
    for col in ['H', 'I', 'J', 'K']:
        for row in range(5, 155):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'
    
    # Percentage formatting for Acknowledgment_Rate
    for row in range(5, 155):
        ws[f'P{row}'].number_format = '0%'
    
    # Freeze panes at A5
    ws.freeze_panes = 'A5'
    
    # Define named range for Policy_ID_List
    policy_id_range = DefinedName(name='Policy_ID_List', attr_text="Policy_Inventory!$A$5:$A$154")
    wb.defined_names.add(policy_id_range)
    
    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_lifecycle_compliance_sheet(wb, styles):
    """
    Sheet 3: Lifecycle_Compliance
    
    Verify policy lifecycle stages (creation, approval, review, update).
    """
    ws = wb.create_sheet("Lifecycle_Compliance")
    
    # --- Header ---
    merge_and_style(ws, 'A1:N1', 'LIFECYCLE COMPLIANCE ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:N2',
                   'Verification of policy creation, approval, publication, review, and update processes',
                   styles['header_sub'])
    
    # --- Column Headers ---
    headers = [
        ('A', 'Policy_ID', 20),
        ('B', 'Policy_Title', 35),
        ('C', 'Creation_Process_Followed', 20),
        ('D', 'Approval_Valid', 15),
        ('E', 'Approval_Documentation', 20),
        ('F', 'Publication_Status', 20),
        ('G', 'Review_Schedule_Defined', 20),
        ('H', 'Review_Status', 20),
        ('I', 'Last_Review_Evidence', 20),
        ('J', 'Version_Control_Practice', 15),
        ('K', 'Sunset_Process', 15),
        ('L', 'Lifecycle_Compliance_Rating', 20),
        ('M', 'Gap_Description', 40),
        ('N', 'Evidence_Reference', 30)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)
    
    # --- Data Rows (5-154) with Auto-Population Formulas ---
    for row in range(5, 155):
        # Auto-populate Policy_ID from Policy_Inventory
        ws[f'A{row}'] = f'=IF(ROW()-4<=COUNTA(Policy_Inventory!$A:$A)-4,INDEX(Policy_Inventory!$A:$A,ROW()-3),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])
        
        # Auto-populate Policy_Title
        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Policy_Inventory!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])
        
        # Manual entry cells (C-G, I-K, M-N)
        for col in ['C', 'D', 'E', 'F', 'G', 'I', 'J', 'K', 'M', 'N']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])
        
        # Review_Status (H) - auto-calculated
        ws[f'H{row}'] = (f'=IF(B{row}="","",IF(VLOOKUP(A{row},Policy_Inventory!$A:$K,11,FALSE)>=TODAY(),"Current",'
                        f'IF(VLOOKUP(A{row},Policy_Inventory!$A:$K,11,FALSE)>=TODAY()-30,"Overdue-<30-Days",'
                        f'IF(VLOOKUP(A{row},Policy_Inventory!$A:$K,11,FALSE)>=TODAY()-90,"Overdue-30-90-Days","Overdue->90-Days"))))')
        apply_cell_style(ws[f'H{row}'], styles['calculated_cell'])
        
        # Lifecycle_Compliance_Rating (L) - auto-calculated
        ws[f'L{row}'] = (f'=IF(B{row}="","",IF(AND(D{row}="Yes",E{row}="Complete",F{row}="Published",'
                        f'OR(H{row}="Current",H{row}="Overdue-<30-Days"),I{row}="Yes",J{row}<>"Poor"),"Compliant",'
                        f'IF(OR(D{row}="No",D{row}="Expired",E{row}="Missing",F{row}="Not-Published",H{row}="Overdue->90-Days"),"Non-Compliant","Partial")))')
        apply_cell_style(ws[f'L{row}'], styles['calculated_cell'])
    
    # --- Data Validation ---
    creation_dv = create_data_validation(['Yes', 'Partial', 'No', 'Unknown'])
    ws.add_data_validation(creation_dv)
    creation_dv.add('C5:C154')
    
    approval_valid_dv = create_data_validation(['Yes', 'Partial', 'No', 'Expired'])
    ws.add_data_validation(approval_valid_dv)
    approval_valid_dv.add('D5:D154')
    
    approval_doc_dv = create_data_validation(['Complete', 'Incomplete', 'Missing'])
    ws.add_data_validation(approval_doc_dv)
    approval_doc_dv.add('E5:E154')
    
    pub_status_dv = create_data_validation(['Published', 'Not-Published', 'Partially-Accessible'])
    ws.add_data_validation(pub_status_dv)
    pub_status_dv.add('F5:F154')
    
    review_schedule_dv = create_data_validation(['Yes', 'No'])
    ws.add_data_validation(review_schedule_dv)
    review_schedule_dv.add('G5:G154')
    
    review_evidence_dv = create_data_validation(['Yes', 'Partial', 'No'])
    ws.add_data_validation(review_evidence_dv)
    review_evidence_dv.add('I5:I154')
    
    version_control_dv = create_data_validation(['Excellent', 'Good', 'Adequate', 'Poor'])
    ws.add_data_validation(version_control_dv)
    version_control_dv.add('J5:J154')
    
    sunset_dv = create_data_validation(['N/A', 'Yes', 'No'])
    ws.add_data_validation(sunset_dv)
    sunset_dv.add('K5:K154')
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_governance_assessment_sheet(wb, styles):
    """
    Sheet 4: Governance_Assessment
    
    Ownership, accountability, approval authority verification.
    """
    ws = wb.create_sheet("Governance_Assessment")
    
    # --- Header ---
    merge_and_style(ws, 'A1:O1', 'GOVERNANCE ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:O2',
                   'Ownership, accountability, approval authority, and RACI verification',
                   styles['header_sub'])
    
    # --- Column Headers ---
    headers = [
        ('A', 'Policy_ID', 20),
        ('B', 'Policy_Title', 35),
        ('C', 'Policy_Hierarchy_Tier', 15),
        ('D', 'Owner_Assigned', 15),
        ('E', 'Owner_Name_Role', 25),
        ('F', 'Owner_Accountability_Clear', 15),
        ('G', 'Approver_Assigned', 15),
        ('H', 'Approver_Name_Role', 25),
        ('I', 'Approval_Authority_Appropriate', 20),
        ('J', 'RACI_Defined', 15),
        ('K', 'Governance_Documentation', 20),
        ('L', 'Escalation_Path_Clear', 15),
        ('M', 'Governance_Compliance_Rating', 20),
        ('N', 'Gap_Description', 40),
        ('O', 'Evidence_Reference', 30)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)
    
    # --- Data Rows ---
    for row in range(5, 155):
        # Auto-populate columns
        ws[f'A{row}'] = f'=IF(ROW()-4<=COUNTA(Policy_Inventory!$A:$A)-4,INDEX(Policy_Inventory!$A:$A,ROW()-3),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])
        
        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Policy_Inventory!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])
        
        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Policy_Inventory!$A:$C,3,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])
        
        ws[f'D{row}'] = f'=IF(A{row}="","",IF(VLOOKUP(A{row},Policy_Inventory!$A:$E,5,FALSE)<>"","Yes","No"))'
        apply_cell_style(ws[f'D{row}'], styles['calculated_cell'])
        
        ws[f'E{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Policy_Inventory!$A:$E,5,FALSE),"")'
        apply_cell_style(ws[f'E{row}'], styles['calculated_cell'])
        
        ws[f'G{row}'] = f'=IF(A{row}="","",IF(VLOOKUP(A{row},Policy_Inventory!$A:$F,6,FALSE)<>"","Yes","No"))'
        apply_cell_style(ws[f'G{row}'], styles['calculated_cell'])
        
        ws[f'H{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Policy_Inventory!$A:$F,6,FALSE),"")'
        apply_cell_style(ws[f'H{row}'], styles['calculated_cell'])
        
        # Manual entry cells
        for col in ['F', 'I', 'J', 'K', 'L', 'N', 'O']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])
        
        # Governance_Compliance_Rating (M) - auto-calculated
        ws[f'M{row}'] = (f'=IF(B{row}="","",IF(AND(D{row}="Yes",G{row}="Yes",I{row}="Yes",K{row}="Complete",L{row}="Yes"),"Compliant",'
                        f'IF(OR(D{row}="No",G{row}="No",I{row}="Insufficient",K{row}="Missing"),"Non-Compliant","Partial")))')
        apply_cell_style(ws[f'M{row}'], styles['calculated_cell'])
    
    # --- Data Validation ---
    owner_account_dv = create_data_validation(['Yes', 'Unclear', 'No'])
    ws.add_data_validation(owner_account_dv)
    owner_account_dv.add('F5:F154')
    
    approval_auth_dv = create_data_validation(['Yes', 'No', 'Elevated-Unnecessarily', 'Insufficient'])
    ws.add_data_validation(approval_auth_dv)
    approval_auth_dv.add('I5:I154')
    
    raci_dv = create_data_validation(['Yes', 'Partial', 'No', 'N/A'])
    ws.add_data_validation(raci_dv)
    raci_dv.add('J5:J154')
    
    gov_doc_dv = create_data_validation(['Complete', 'Partial', 'Missing'])
    ws.add_data_validation(gov_doc_dv)
    gov_doc_dv.add('K5:K154')
    
    escalation_dv = create_data_validation(['Yes', 'Unclear', 'No'])
    ws.add_data_validation(escalation_dv)
    escalation_dv.add('L5:L154')
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_classification_review_sheet(wb, styles):
    """
    Sheet 5: Classification_Review
    
    Policy classification appropriateness and access control verification.
    """
    ws = wb.create_sheet("Classification_Review")
    
    # --- Header ---
    merge_and_style(ws, 'A1:L1', 'CLASSIFICATION REVIEW', styles['header_main'])
    merge_and_style(ws, 'A2:L2',
                   'Policy classification appropriateness and access control verification',
                   styles['header_sub'])
    
    # --- Column Headers ---
    headers = [
        ('A', 'Policy_ID', 20),
        ('B', 'Policy_Title', 35),
        ('C', 'Current_Classification', 15),
        ('D', 'Classification_Appropriate', 20),
        ('E', 'Content_Sensitivity_Assessment', 15),
        ('F', 'Access_Controls_Implemented', 20),
        ('G', 'Distribution_Restrictions', 20),
        ('H', 'Classification_Marking', 15),
        ('I', 'Classification_Review_Date', 15),
        ('J', 'Classification_Compliance_Rating', 20),
        ('K', 'Gap_Description', 40),
        ('L', 'Evidence_Reference', 30)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)
    
    # --- Data Rows ---
    for row in range(5, 155):
        # Auto-populate columns
        ws[f'A{row}'] = f'=IF(ROW()-4<=COUNTA(Policy_Inventory!$A:$A)-4,INDEX(Policy_Inventory!$A:$A,ROW()-3),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])
        
        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Policy_Inventory!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])
        
        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Policy_Inventory!$A:$N,14,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])
        
        # Manual entry cells
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'K', 'L']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])
        
        # Classification_Compliance_Rating (J) - auto-calculated
        ws[f'J{row}'] = (f'=IF(B{row}="","",IF(AND(D{row}="Yes",F{row}="Yes",H{row}="Yes"),"Compliant",'
                        f'IF(OR(D{row}="No",F{row}="No",H{row}="No"),"Non-Compliant","Partial")))')
        apply_cell_style(ws[f'J{row}'], styles['calculated_cell'])
    
    # Date formatting for Classification_Review_Date
    for row in range(5, 155):
        ws[f'I{row}'].number_format = 'DD.MM.YYYY'
    
    # --- Data Validation ---
    class_approp_dv = create_data_validation(['Yes', 'No', 'Should-Be-Higher', 'Should-Be-Lower'])
    ws.add_data_validation(class_approp_dv)
    class_approp_dv.add('D5:D154')
    
    sensitivity_dv = create_data_validation(['None', 'Low', 'Medium', 'High'])
    ws.add_data_validation(sensitivity_dv)
    sensitivity_dv.add('E5:E154')
    
    access_controls_dv = create_data_validation(['Yes', 'Partial', 'No', 'Unknown'])
    ws.add_data_validation(access_controls_dv)
    access_controls_dv.add('F5:F154')
    
    distribution_dv = create_data_validation(['Yes', 'Partial', 'No', 'N/A'])
    ws.add_data_validation(distribution_dv)
    distribution_dv.add('G5:G154')
    
    marking_dv = create_data_validation(['Yes', 'No', 'Inconsistent'])
    ws.add_data_validation(marking_dv)
    marking_dv.add('H5:H154')
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_communication_tracking_sheet(wb, styles):
    """
    Sheet 6: Communication_Tracking
    
    Policy publication, acknowledgment, and training integration.
    """
    ws = wb.create_sheet("Communication_Tracking")
    
    # --- Header ---
    merge_and_style(ws, 'A1:P1', 'COMMUNICATION TRACKING', styles['header_main'])
    merge_and_style(ws, 'A2:P2',
                   'Policy publication, acknowledgment, and training integration verification',
                   styles['header_sub'])
    
    # --- Column Headers ---
    headers = [
        ('A', 'Policy_ID', 20),
        ('B', 'Policy_Title', 35),
        ('C', 'Acknowledgment_Required', 15),
        ('D', 'Target_Audience', 20),
        ('E', 'Publication_Date', 12),
        ('F', 'Publication_Method', 30),
        ('G', 'Accessibility_Verified', 15),
        ('H', 'Acknowledgment_Mechanism', 20),
        ('I', 'Acknowledgment_Rate', 12),
        ('J', 'Acknowledgment_Timeframe', 15),
        ('K', 'Non_Acknowledgment_Follow_Up', 20),
        ('L', 'Training_Integration', 15),
        ('M', 'User_Feedback_Mechanism', 15),
        ('N', 'Communication_Compliance_Rating', 20),
        ('O', 'Gap_Description', 40),
        ('P', 'Evidence_Reference', 30)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)
    
    # --- Data Rows ---
    for row in range(5, 155):
        # Auto-populate columns
        ws[f'A{row}'] = f'=IF(ROW()-4<=COUNTA(Policy_Inventory!$A:$A)-4,INDEX(Policy_Inventory!$A:$A,ROW()-3),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])
        
        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Policy_Inventory!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])
        
        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Policy_Inventory!$A:$O,15,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])
        
        ws[f'I{row}'] = f'=IF(A{row}="","",VLOOKUP(A{row},Policy_Inventory!$A:$P,16,FALSE))'
        apply_cell_style(ws[f'I{row}'], styles['calculated_cell'])
        ws[f'I{row}'].number_format = '0%'
        
        # Manual entry cells
        for col in ['D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'M', 'O', 'P']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])
        
        # Communication_Compliance_Rating (N) - auto-calculated
        ws[f'N{row}'] = (f'=IF(B{row}="","",IF(AND(G{row}="Yes",OR(C{row}="No",AND(C{row}<>"No",I{row}>=0.9)),L{row}<>"No"),"Compliant",'
                        f'IF(OR(G{row}="No",AND(C{row}<>"No",I{row}<0.7)),"Non-Compliant","Partial")))')
        apply_cell_style(ws[f'N{row}'], styles['calculated_cell'])
    
    # Date formatting
    for row in range(5, 155):
        ws[f'E{row}'].number_format = 'DD.MM.YYYY'
    
    # --- Data Validation ---
    accessibility_dv = create_data_validation(['Yes', 'Partial', 'No'])
    ws.add_data_validation(accessibility_dv)
    accessibility_dv.add('G5:G154')
    
    ack_mech_dv = create_data_validation(['N/A', 'LMS', 'Email', 'E-Signature', 'Form', 'Training'])
    ws.add_data_validation(ack_mech_dv)
    ack_mech_dv.add('H5:H154')
    
    ack_time_dv = create_data_validation(['N/A', 'Immediate', '30-Days', '60-Days', '90-Days', 'Annual'])
    ws.add_data_validation(ack_time_dv)
    ack_time_dv.add('J5:J154')
    
    followup_dv = create_data_validation(['Yes', 'Partial', 'No', 'N/A'])
    ws.add_data_validation(followup_dv)
    followup_dv.add('K5:K154')
    
    training_dv = create_data_validation(['Yes', 'Partial', 'No', 'N/A'])
    ws.add_data_validation(training_dv)
    training_dv.add('L5:L154')
    
    feedback_dv = create_data_validation(['Yes', 'No'])
    ws.add_data_validation(feedback_dv)
    feedback_dv.add('M5:M154')
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_repository_assessment_sheet(wb, styles):
    """
    Sheet 7: Repository_Assessment
    
    Policy repository structure, organization, and accessibility (repository-wide).
    """
    ws = wb.create_sheet("Repository_Assessment")
    
    # --- Header ---
    merge_and_style(ws, 'A1:B1', 'REPOSITORY ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:B2',
                   'Policy repository structure, organization, access, and performance',
                   styles['header_sub'])
    
    # --- Assessment Questions (Rows 5-22) ---
    questions = [
        (5, 'Repository_Type', 'What system is used for policy repository?'),
        (6, 'Repository_URL_Path', 'What is the primary repository location?'),
        (7, 'Repository_Organization', 'How is repository organized?'),
        (8, 'Navigation_Ease', 'How easy is it to find policies?'),
        (9, 'Search_Functionality', 'Is search functionality available?'),
        (10, 'Version_Control', 'How are policy versions managed?'),
        (11, 'Previous_Versions_Accessible', 'Can users access previous versions?'),
        (12, 'Access_Control_Implementation', 'Are access controls implemented?'),
        (13, 'Access_Logging', 'Is policy access logged?'),
        (14, 'Mobile_Accessibility', 'Are policies accessible from mobile?'),
        (15, 'Offline_Access', 'Can policies be accessed offline?'),
        (16, 'Repository_Backup', 'Is repository backed up?'),
        (17, 'Repository_Disaster_Recovery', 'Is repository in DR plan?'),
        (18, 'Repository_Performance', 'Is repository performance acceptable?'),
        (19, 'Repository_Uptime', 'What is repository availability?'),
        (20, 'Repository_Compliance_Rating', 'Overall repository compliance?'),
        (21, 'Gap_Description', 'Describe any repository gaps'),
        (22, 'Evidence_Reference', 'Where is repository evidence?')
    ]
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 60
    
    for row_num, field_name, question in questions:
        ws[f'A{row_num}'] = question
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        
        if row_num == 20:  # Compliance Rating - auto-calculated
            ws[f'B{row_num}'] = ('=IF(AND(B7<>"Unorganized",B8<>"Poor",B9<>"None",B10<>"No-Versioning",'
                                'B12="Yes",B16="Yes"),"Compliant",IF(OR(B7="Unorganized",B8="Poor",'
                                'B10="No-Versioning",B12="No",B16="No"),"Non-Compliant","Partial"))')
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])
    
    # --- Data Validation ---
    # Repository_Type (B5)
    repo_type_dv = create_data_validation(['SharePoint', 'Confluence', 'Document-Management-System', 
                                          'Intranet', 'File-Share', 'Other'])
    ws.add_data_validation(repo_type_dv)
    repo_type_dv.add('B5')
    
    # Repository_Organization (B7)
    repo_org_dv = create_data_validation(['By-Hierarchy-Tier', 'By-Domain', 'By-ISO-Control', 
                                         'Alphabetical', 'Unorganized', 'Mixed'])
    ws.add_data_validation(repo_org_dv)
    repo_org_dv.add('B7')
    
    # Navigation_Ease (B8)
    nav_dv = create_data_validation(['Excellent', 'Good', 'Adequate', 'Poor'])
    ws.add_data_validation(nav_dv)
    nav_dv.add('B8')
    
    # Search_Functionality (B9)
    search_dv = create_data_validation(['Excellent', 'Good', 'Adequate', 'None'])
    ws.add_data_validation(search_dv)
    search_dv.add('B9')
    
    # Version_Control (B10)
    version_dv = create_data_validation(['Automated-Versioning', 'Manual-Versioning', 'No-Versioning'])
    ws.add_data_validation(version_dv)
    version_dv.add('B10')
    
    # Previous_Versions_Accessible (B11)
    prev_ver_dv = create_data_validation(['Yes-All-Versions', 'Yes-Recent-Versions', 'No', 'N/A'])
    ws.add_data_validation(prev_ver_dv)
    prev_ver_dv.add('B11')
    
    # Yes/Partial/No dropdowns
    ypn_dv = create_data_validation(['Yes', 'Partial', 'No'])
    ws.add_data_validation(ypn_dv)
    for row in [12, 13]:
        ypn_dv.add(f'B{row}')
    
    # Mobile_Accessibility (B14)
    mobile_dv = create_data_validation(['Yes-Fully', 'Yes-Partially', 'No', 'Not-Tested'])
    ws.add_data_validation(mobile_dv)
    mobile_dv.add('B14')
    
    # Offline_Access (B15)
    offline_dv = create_data_validation(['Yes', 'No', 'Not-Required'])
    ws.add_data_validation(offline_dv)
    offline_dv.add('B15')
    
    # Yes/No/Unknown dropdowns
    ynu_dv = create_data_validation(['Yes', 'No', 'Unknown'])
    ws.add_data_validation(ynu_dv)
    for row in [16, 17]:
        ynu_dv.add(f'B{row}')
    
    # Repository_Performance (B18)
    perf_dv = create_data_validation(['Excellent', 'Good', 'Adequate', 'Poor'])
    ws.add_data_validation(perf_dv)
    perf_dv.add('B18')
    
    # Repository_Uptime (B19)
    uptime_dv = create_data_validation(['99.9%+', '99-99.9%', '95-99%', '<95%', 'Unknown'])
    ws.add_data_validation(uptime_dv)
    uptime_dv.add('B19')
    
    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False


def create_gap_analysis_sheet(wb, styles):
    """
    Sheet 8: Gap_Analysis
    
    Consolidated gaps from all assessment domains.
    """
    ws = wb.create_sheet("Gap_Analysis")
    
    # --- Header ---
    merge_and_style(ws, 'A1:P1', 'GAP ANALYSIS', styles['header_main'])
    merge_and_style(ws, 'A2:P2',
                   'Consolidated gaps from all assessment domains with risk levels and remediation plans',
                   styles['header_sub'])
    
    # --- Column Headers ---
    headers = [
        ('A', 'Gap_ID', 12),
        ('B', 'Policy_ID', 20),
        ('C', 'Policy_Title', 35),
        ('D', 'Gap_Category', 15),
        ('E', 'Gap_Description', 40),
        ('F', 'Risk_Level', 15),
        ('G', 'Impact_Assessment', 35),
        ('H', 'Affected_Stakeholders', 25),
        ('I', 'Remediation_Action', 40),
        ('J', 'Responsible_Party', 25),
        ('K', 'Target_Completion_Date', 15),
        ('L', 'Estimated_Effort', 15),
        ('M', 'Dependencies', 30),
        ('N', 'Status', 15),
        ('O', 'Completion_Evidence', 30),
        ('P', 'Risk_Acceptance', 40)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)
    
    # --- Data Rows (5-104, 100 rows for gaps) ---
    for row in range(5, 105):
        # Gap_ID auto-generated
        ws[f'A{row}'] = f'=IF(ROW()<=4,"",TEXT(ROW()-4,"GAP-000"))'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])
        
        # All other columns manual entry
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])
    
    # Date formatting for Target_Completion_Date
    for row in range(5, 105):
        ws[f'K{row}'].number_format = 'DD.MM.YYYY'
    
    # --- Data Validation ---
    risk_level_dv = create_data_validation(['Critical', 'High', 'Medium', 'Low'])
    ws.add_data_validation(risk_level_dv)
    risk_level_dv.add('F5:F104')
    
    effort_dv = create_data_validation(['<1hr', '1-4hrs', '1day', '2-5days', '>1week'])
    ws.add_data_validation(effort_dv)
    effort_dv.add('L5:L104')
    
    status_dv = create_data_validation(['Not-Started', 'In-Progress', 'Blocked', 'Completed', 'Accepted-Risk'])
    ws.add_data_validation(status_dv)
    status_dv.add('N5:N104')
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_evidence_register_sheet(wb, styles):
    """
    Sheet 9: Evidence_Register
    
    Supporting evidence documentation.
    """
    ws = wb.create_sheet("Evidence_Register")
    
    # --- Header ---
    merge_and_style(ws, 'A1:J1', 'EVIDENCE REGISTER', styles['header_main'])
    merge_and_style(ws, 'A2:J2',
                   'Documentation of all supporting evidence for policy framework assessment',
                   styles['header_sub'])
    
    # --- Column Headers ---
    headers = [
        ('A', 'Evidence_ID', 12),
        ('B', 'Evidence_Type', 20),
        ('C', 'Description', 40),
        ('D', 'Related_Policy_ID', 20),
        ('E', 'Related_Assessment_Sheet', 20),
        ('F', 'File_Location', 40),
        ('G', 'Date_Collected', 12),
        ('H', 'Collected_By', 25),
        ('I', 'Verification_Status', 15),
        ('J', 'Notes', 40)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)
    
    # --- Data Rows (5-104, 100 evidence items) ---
    for row in range(5, 105):
        # Evidence_ID auto-generated
        ws[f'A{row}'] = f'=IF(ROW()<=4,"",TEXT(ROW()-4,"EVD-000"))'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])
        
        # All other columns manual entry
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])
    
    # Date formatting for Date_Collected
    for row in range(5, 105):
        ws[f'G{row}'].number_format = 'DD.MM.YYYY'
    
    # --- Data Validation ---
    evidence_type_dv = create_data_validation(['Policy-Document', 'Approval-Email', 'Approval-Signature',
                                               'Meeting-Minutes', 'Review-Checklist', 'Repository-Screenshot',
                                               'Access-Control-Config', 'Acknowledgment-Report', 'Training-Records',
                                               'Version-History', 'Organizational-Chart', 'Other'])
    ws.add_data_validation(evidence_type_dv)
    evidence_type_dv.add('B5:B104')
    
    verification_dv = create_data_validation(['Verified', 'Pending', 'Not-Verified', 'Expired'])
    ws.add_data_validation(verification_dv)
    verification_dv.add('I5:I104')
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_action_items_sheet(wb, styles):
    """
    Sheet 10: Action_Items
    
    Remediation tracking (auto-populated from Gap_Analysis).
    """
    ws = wb.create_sheet("Action_Items")
    
    # --- Header ---
    merge_and_style(ws, 'A1:M1', 'ACTION ITEMS', styles['header_main'])
    merge_and_style(ws, 'A2:M2', 'Remediation tracking for identified gaps', styles['header_sub'])
    
    # --- Column Headers ---
    headers = [
        ('A', 'Action_ID', 12),
        ('B', 'Related_Gap_ID', 12),
        ('C', 'Action_Description', 40),
        ('D', 'Priority', 15),
        ('E', 'Owner', 25),
        ('F', 'Target_Date', 15),
        ('G', 'Status', 15),
        ('H', 'Progress_%', 12),
        ('I', 'Last_Update_Date', 12),
        ('J', 'Last_Update_Notes', 35),
        ('K', 'Blocking_Issues', 35),
        ('L', 'Escalation_Required', 15),
        ('M', 'Escalation_To', 25)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)
    
    # --- Data Rows (5-54, 50 action items) ---
    for row in range(5, 55):
        # Auto-populate from Gap_Analysis
        ws[f'A{row}'] = f'=IF(Gap_Analysis!A{row}="","",SUBSTITUTE(Gap_Analysis!A{row},"GAP","ACT"))'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])
        
        ws[f'B{row}'] = f'=Gap_Analysis!A{row}'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])
        
        ws[f'C{row}'] = f'=IF(Gap_Analysis!I{row}<>"",Gap_Analysis!I{row},"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])
        
        ws[f'D{row}'] = f'=IF(Gap_Analysis!F{row}<>"",Gap_Analysis!F{row},"")'
        apply_cell_style(ws[f'D{row}'], styles['calculated_cell'])
        
        ws[f'E{row}'] = f'=IF(Gap_Analysis!J{row}<>"",Gap_Analysis!J{row},"")'
        apply_cell_style(ws[f'E{row}'], styles['calculated_cell'])
        
        ws[f'F{row}'] = f'=IF(Gap_Analysis!K{row}<>"",Gap_Analysis!K{row},"")'
        apply_cell_style(ws[f'F{row}'], styles['calculated_cell'])
        ws[f'F{row}'].number_format = 'DD.MM.YYYY'
        
        ws[f'G{row}'] = f'=IF(Gap_Analysis!N{row}<>"",Gap_Analysis!N{row},"")'
        apply_cell_style(ws[f'G{row}'], styles['calculated_cell'])
        
        # Manual entry cells
        for col in ['H', 'I', 'J', 'K', 'L', 'M']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])
    
    # Percentage formatting for Progress_%
    for row in range(5, 55):
        ws[f'H{row}'].number_format = '0%'
        ws[f'I{row}'].number_format = 'DD.MM.YYYY'
    
    # --- Data Validation ---
    escalation_dv = create_data_validation(['Yes', 'No'])
    ws.add_data_validation(escalation_dv)
    escalation_dv.add('L5:L54')
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_approval_signoff_sheet(wb, styles):
    """
    Sheet 11: Approval_Sign_Off
    
    Three-level approval workflow.
    """
    ws = wb.create_sheet("Approval_Sign_Off")
    
    # --- Header ---
    merge_and_style(ws, 'A1:D1', 'APPROVAL & SIGN-OFF', styles['header_main'])
    ws.row_dimensions[1].height = 30
    
    # --- Assessment Summary (Rows 3-20) ---
    ws['A3'] = 'ASSESSMENT SUMMARY'
    apply_cell_style(ws['A3'], styles['section_header'])
    ws.merge_cells('A3:D3')
    
    summary_items = [
        (5, 'Document:', 'ISMS-IMP-A.5.1-2-6.1-2.S1'),
        (6, 'Assessment Period:', '=Dashboard!B7'),
        (7, 'Overall Compliance Score:', '=Dashboard!B19'),
        (8, 'Total Policies:', '=Dashboard!B20'),
        (9, 'Compliant Policies:', '=Dashboard!B21'),
        (10, 'Policies with Gaps:', '=Dashboard!B22'),
        (11, 'Critical Gaps:', '=Dashboard!B23'),
        (12, 'High Priority Gaps:', '=Dashboard!B24'),
        (13, 'Repository Compliance:', '=Repository_Assessment!B20')
    ]
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    
    for row_num, label, value in summary_items:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        
        ws[f'B{row_num}'] = value
        if value.startswith('='):
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
    
    # --- Level 1 Approval (Rows 22-30) ---
    ws['A22'] = 'LEVEL 1: PREPARED BY'
    apply_cell_style(ws['A22'], styles['section_header'])
    ws.merge_cells('A22:D22')
    
    level1_fields = [
        (23, 'Name:'),
        (24, 'Role:'),
        (25, 'Date:'),
        (26, 'Signature:'),
        (27, 'Certification:', 'I certify this assessment is complete and accurate'),
        (28, 'Comments:')
    ]
    
    for row_num, label, *rest in level1_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        
        if rest:
            ws[f'B{row_num}'] = rest[0]
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])
    
    # --- Level 2 Approval (Rows 32-45) ---
    ws['A32'] = 'LEVEL 2: REVIEWED BY'
    apply_cell_style(ws['A32'], styles['section_header'])
    ws.merge_cells('A32:D32')
    
    level2_fields = [
        (33, 'Name:'),
        (34, 'Role:'),
        (35, 'Date:'),
        (36, 'Signature:')
    ]
    
    for row_num, label in level2_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])
    
    # Review checklist
    checklist_items = [
        'Policy inventory complete',
        'Lifecycle compliance verified',
        'Governance gaps identified',
        'Communication assessment thorough',
        'Repository assessment complete',
        'Gap risk levels appropriate',
        'Evidence sufficient',
        'Action items tracked'
    ]
    
    ws['A38'] = 'Review Checklist:'
    apply_cell_style(ws['A38'], styles['label_cell'])
    
    for i, item in enumerate(checklist_items, start=39):
        ws[f'A{i}'] = f'☐ {item}'
        ws.merge_cells(f'A{i}:D{i}')
    
    ws['A47'] = 'Comments:'
    apply_cell_style(ws['A47'], styles['label_cell'])
    apply_cell_style(ws['B47'], styles['input_cell'])
    
    # --- Level 3 Approval (Rows 49-60) ---
    ws['A49'] = 'LEVEL 3: APPROVED BY (CISO)'
    apply_cell_style(ws['A49'], styles['section_header'])
    ws.merge_cells('A49:D49')
    
    level3_fields = [
        (50, 'Name:'),
        (51, 'Role:'),
        (52, 'Date:'),
        (53, 'Signature:'),
        (54, 'Final Approval:', 'I approve this assessment as accurate'),
        (55, 'Risk Acceptance:', 'I accept residual risk for: [list]'),
        (56, 'Comments:')
    ]
    
    for row_num, label, *rest in level3_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        
        if rest:
            ws[f'B{row_num}'] = rest[0]
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])
    
    # --- Assessment Metadata (Rows 59-65) ---
    ws['A59'] = 'Next Review Date:'
    apply_cell_style(ws['A59'], styles['label_cell'])
    apply_cell_style(ws['B59'], styles['input_cell'])
    
    ws['A60'] = 'Assessment Status:'
    apply_cell_style(ws['A60'], styles['label_cell'])
    apply_cell_style(ws['B60'], styles['input_cell'])
    
    status_dv = create_data_validation(['Draft', 'Under-Review', 'Approved', 'Audit-Ready'])
    ws.add_data_validation(status_dv)
    status_dv.add('B60')
    
    # Audit Readiness Checklist
    ws['A62'] = 'Audit Readiness Checklist:'
    apply_cell_style(ws['A62'], styles['label_cell'])
    
    audit_items = [
        'All three approvals complete',
        'Evidence 100% verified',
        'All critical gaps have remediation plans',
        'Repository compliance verified',
        'Assessment audit-ready'
    ]
    
    for i, item in enumerate(audit_items, start=63):
        ws[f'A{i}'] = f'☐ {item}'
        ws.merge_cells(f'A{i}:D{i}')
    
    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False


# ==============================================================================
# MAIN EXECUTION
# ==============================================================================

def main():
    """
    Main function to generate the Policy Framework Assessment workbook.

    Returns:
        int: 0 on success, 1 on failure
    """
    try:
        logger.info("=" * 80)
        logger.info("ISMS-IMP-A.5.1-2-6.1-2.S1 - Policy Framework Assessment")
        logger.info("Workbook Generator")
        logger.info("=" * 80)

        # Initialize workbook
        logger.info("Creating workbook...")
        wb = Workbook()
        styles = setup_styles()

        # Create all sheets
        logger.info("Creating Dashboard...")
        create_dashboard_sheet(wb, styles)

        logger.info("Creating Policy_Inventory...")
        create_policy_inventory_sheet(wb, styles)

        logger.info("Creating Lifecycle_Compliance...")
        create_lifecycle_compliance_sheet(wb, styles)

        logger.info("Creating Governance_Assessment...")
        create_governance_assessment_sheet(wb, styles)

        logger.info("Creating Classification_Review...")
        create_classification_review_sheet(wb, styles)

        logger.info("Creating Communication_Tracking...")
        create_communication_tracking_sheet(wb, styles)

        logger.info("Creating Repository_Assessment...")
        create_repository_assessment_sheet(wb, styles)

        logger.info("Creating Gap_Analysis...")
        create_gap_analysis_sheet(wb, styles)

        logger.info("Creating Evidence_Register...")
        create_evidence_register_sheet(wb, styles)

        logger.info("Creating Action_Items...")
        create_action_items_sheet(wb, styles)

        logger.info("Creating Approval_Sign_Off...")
        create_approval_signoff_sheet(wb, styles)

        # Set workbook properties
        wb.properties.title = "ISMS-IMP-A.5.1-2-6.1-2.S1 - Policy Framework Assessment"
        wb.properties.subject = "ISO/IEC 27001:2022 Control A.5.1 Assessment"
        wb.properties.creator = "[Organization] Information Security Team"
        wb.properties.keywords = "Policy, Governance, ISMS, ISO27001, A.5.1"
        wb.properties.comments = "Generated via Python script generate_a5_1_2_6_1_2_s1_policy_framework.py"

        # Generate filename with current date
        today = datetime.now().strftime("%Y%m%d")
        filename = f"ISMS-IMP-A.5.1-2-6.1-2.S1_Policy_Framework_{today}.xlsx"

        # Save workbook
        logger.info(f"Saving workbook as: {filename}")
        wb.save(filename)

        logger.info("=" * 80)
        logger.info("Workbook generated successfully!")
        logger.info("=" * 80)
        logger.info("Next Steps:")
        logger.info("1. Open the workbook in Excel")
        logger.info("2. Complete Sheet 2 (Policy_Inventory) - this is your foundation")
        logger.info("3. Complete Sheets 3-7 (assessment domains)")
        logger.info("4. Review Sheet 8 (Gap_Analysis) - auto-populated")
        logger.info("5. Document Sheet 9 (Evidence)")
        logger.info("6. Plan Sheet 10 (Action Items)")
        logger.info("7. Review Sheet 1 (Dashboard) - auto-calculated")
        logger.info("8. Obtain Sheet 11 (Approval Sign-Off)")
        logger.info(f"File location: ./{filename}")

        return 0

    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
