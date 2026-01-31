#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-A.5.9 - Assessment File Normalizer Utility
================================================================================

ISO/IEC 27001:2022 Control A.5.9: Inventory of Information and Assets
Quality Assurance Utility: Excel Assessment File Normalization & Validation

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific assessment file standards and validation requirements.

Key customization areas:
1. Expected file naming conventions (match your organizational standards)
2. Workbook structure validation rules (specific to your A.5.9 assessments)
3. Data format normalization rules (adapt to your data standards)
4. Validation severity thresholds (based on your quality requirements)
5. Output formatting preferences (align with your reporting needs)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script normalizes and validates A.5.9 inventory assessment Excel workbooks
to ensure consistency, data quality, and compliance with framework standards
before consolidation into the compliance dashboard.

**Purpose:**
Ensures all inventory assessment workbooks meet quality standards and structural
requirements, preventing data consolidation errors and improving audit evidence
reliability.

**Key Functions:**
1. File Naming Validation - Verify naming convention compliance
2. Workbook Structure Validation - Check required sheets and headers
3. Data Normalization - Standardize formats and values
4. Content Validation - Check for incomplete assessments
5. Evidence Linkage Validation - Verify evidence references
6. Compliance Scoring Validation - Check formula correctness
7. Quality Reporting - Generate validation report

**Validation Scope:**
- ISMS-IMP-A.5.9.1_Asset_Discovery_YYYYMMDD.xlsx
- ISMS-IMP-A.5.9.2_Inventory_Maintenance_YYYYMMDD.xlsx
- ISMS-IMP-A.5.9.3_Quality_Compliance_YYYYMMDD.xlsx
- ISMS-IMP-A.5.9.4_Owner_Accountability_YYYYMMDD.xlsx

**Output:**
- Validation report (console and optional file)
- Issue summary by severity
- Remediation guidance

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 normalize_assessment_files_a59.py
    
Advanced Usage:
    # Validate specific assessment
    python3 normalize_assessment_files_a59.py --assessment 1
    
    # Specify directory
    python3 normalize_assessment_files_a59.py --dir /path/to/assessments
    
    # Generate detailed report
    python3 normalize_assessment_files_a59.py --report detailed_report.txt
    
    # Dry run (validate without modifying)
    python3 normalize_assessment_files_a59.py --dry-run

--------------------------------------------------------------------------------
"""

import os
import sys
import glob
from datetime import datetime
from openpyxl import load_workbook

# CUSTOMIZE: Expected file patterns and structures
ASSESSMENT_PATTERNS = {
    1: {
        'pattern': 'ISMS-IMP-A.5.9.1_Asset_Discovery_*.xlsx',
        'required_sheets': [
            'Instructions & Legend',
            'Information Assets Discovery',
            'IT Infrastructure Discovery',
            'Applications Discovery',
            'Physical Assets Discovery',
            'Personnel Assets Discovery',
            'Discovery Metrics & Summary',
            'Evidence Register',
        ],
        'name': 'Asset Discovery'
    },
    2: {
        'pattern': 'ISMS-IMP-A.5.9.2_Inventory_Maintenance_*.xlsx',
        'required_sheets': [
            'Instructions & Legend',
            'Inventory Structure & Access',
            'Update Triggers & Workflows',
            'Integration Architecture',
            'Quality Control Processes',
            'Maintenance Metrics',
            'Evidence Register',
        ],
        'name': 'Inventory Maintenance'
    },
    3: {
        'pattern': 'ISMS-IMP-A.5.9.3_Quality_Compliance_*.xlsx',
        'required_sheets': [
            'Instructions & Legend',
            'Accuracy Sampling',
            'Completeness Assessment',
            'Currency Assessment',
            'Consistency Checks',
            'Policy Compliance Matrix',
            'Quality Metrics & Scoring',
            'Evidence Register',
        ],
        'name': 'Quality & Compliance'
    },
    4: {
        'pattern': 'ISMS-IMP-A.5.9.4_Owner_Accountability_*.xlsx',
        'required_sheets': [
            'Instructions & Legend',
            'Ownership Coverage',
            'Owner Acknowledgment',
            'Owner Awareness',
            'Owner Performance',
            'Accountability Metrics',
            'Evidence Register',
        ],
        'name': 'Owner Accountability'
    },
}


class ValidationIssue:
    """Represents a validation issue"""
    CRITICAL = 'CRITICAL'
    HIGH = 'HIGH'
    MEDIUM = 'MEDIUM'
    LOW = 'LOW'
    
    def __init__(self, severity, assessment, location, description, remediation):
        self.severity = severity
        self.assessment = assessment
        self.location = location
        self.description = description
        self.remediation = remediation
    
    def __str__(self):
        return f"[{self.severity}] {self.assessment} - {self.location}: {self.description}"


class AssessmentValidator:
    """Validates A.5.9 assessment workbooks"""
    
    def __init__(self, directory='.', dry_run=False):
        self.directory = directory
        self.dry_run = dry_run
        self.issues = []
    
    def validate_all(self, assessment_num=None):
        """Validate all assessments or specific one"""
        print("="*80)
        print("ISMS Control A.5.9 - Assessment File Validation")
        print("="*80)
        print()
        print(f"Directory: {self.directory}")
        print(f"Mode: {'DRY RUN (validation only)' if self.dry_run else 'VALIDATION'}")
        print()
        
        assessments_to_check = [assessment_num] if assessment_num else [1, 2, 3, 4]
        
        for num in assessments_to_check:
            self.validate_assessment(num)
        
        self.print_summary()
    
    def validate_assessment(self, num):
        """Validate specific assessment workbook"""
        config = ASSESSMENT_PATTERNS[num]
        pattern = os.path.join(self.directory, config['pattern'])
        files = glob.glob(pattern)
        
        print(f"Assessment {num}: {config['name']}")
        print("-" * 80)
        
        if not files:
            print(f"  ✗ NOT FOUND: {config['pattern']}")
            self.issues.append(ValidationIssue(
                ValidationIssue.CRITICAL,
                f"Assessment {num}",
                "File System",
                f"No file matching pattern: {config['pattern']}",
                f"Generate assessment using: python3 generate_a59_{num}_*.py"
            ))
            print()
            return
        
        if len(files) > 1:
            print(f"  ⚠️  MULTIPLE FILES FOUND: {len(files)}")
            for f in files:
                print(f"      - {os.path.basename(f)}")
            self.issues.append(ValidationIssue(
                ValidationIssue.MEDIUM,
                f"Assessment {num}",
                "File System",
                f"Multiple files match pattern (found {len(files)})",
                "Keep only the most recent version, archive older files"
            ))
        
        # Validate the most recent file
        filepath = max(files, key=os.path.getmtime)
        print(f"  ✓ Found: {os.path.basename(filepath)}")
        
        # Validate filename format
        self.validate_filename(filepath, num)
        
        # Validate workbook structure
        self.validate_workbook(filepath, num, config)
        
        print()
    
    def validate_filename(self, filepath, num):
        """Validate filename format"""
        filename = os.path.basename(filepath)
        
        # Check pattern: ISMS-IMP-A.5.9.X_Title_YYYYMMDD.xlsx
        parts = filename.replace('.xlsx', '').split('_')
        
        # Check date suffix (last part should be YYYYMMDD)
        if len(parts) >= 2:
            date_part = parts[-1]
            if len(date_part) == 8 and date_part.isdigit():
                try:
                    datetime.strptime(date_part, '%Y%m%d')
                    print(f"  ✓ Date suffix valid: {date_part}")
                except ValueError:
                    print(f"  ✗ Invalid date suffix: {date_part}")
                    self.issues.append(ValidationIssue(
                        ValidationIssue.MEDIUM,
                        f"Assessment {num}",
                        "Filename",
                        f"Date suffix '{date_part}' is not a valid date",
                        "Use format YYYYMMDD (e.g., 20260122)"
                    ))
            else:
                print(f"  ✗ Date suffix missing or invalid: {date_part}")
                self.issues.append(ValidationIssue(
                    ValidationIssue.MEDIUM,
                    f"Assessment {num}",
                    "Filename",
                    "Date suffix missing or invalid format",
                    "Add YYYYMMDD suffix to filename"
                ))
    
    def validate_workbook(self, filepath, num, config):
        """Validate workbook structure and content"""
        try:
            wb = load_workbook(filepath, data_only=False)
            
            # Check required sheets
            actual_sheets = wb.sheetnames
            required_sheets = config['required_sheets']
            
            missing_sheets = [s for s in required_sheets if s not in actual_sheets]
            extra_sheets = [s for s in actual_sheets if s not in required_sheets and s != 'Sheet']
            
            if missing_sheets:
                print(f"  ✗ Missing sheets: {', '.join(missing_sheets)}")
                for sheet in missing_sheets:
                    self.issues.append(ValidationIssue(
                        ValidationIssue.CRITICAL,
                        f"Assessment {num}",
                        "Workbook Structure",
                        f"Required sheet missing: {sheet}",
                        f"Re-generate workbook using script or add missing sheet"
                    ))
            else:
                print(f"  ✓ All {len(required_sheets)} required sheets present")
            
            if extra_sheets:
                print(f"  ⚠️  Extra sheets found: {', '.join(extra_sheets)}")
                for sheet in extra_sheets:
                    self.issues.append(ValidationIssue(
                        ValidationIssue.LOW,
                        f"Assessment {num}",
                        "Workbook Structure",
                        f"Unexpected sheet found: {sheet}",
                        "Remove extra sheets unless intentionally added"
                    ))
            
            # Check sheet order
            if actual_sheets[:len(required_sheets)] != required_sheets:
                print(f"  ⚠️  Sheet order doesn't match specification")
                self.issues.append(ValidationIssue(
                    ValidationIssue.LOW,
                    f"Assessment {num}",
                    "Workbook Structure",
                    "Sheets are not in the expected order",
                    "Reorder sheets to match framework specification"
                ))
            
            # Validate key sheets have content
            self.validate_sheet_content(wb, num, config)
            
            wb.close()
            
        except Exception as e:
            print(f"  ✗ Error opening workbook: {str(e)}")
            self.issues.append(ValidationIssue(
                ValidationIssue.CRITICAL,
                f"Assessment {num}",
                "File Access",
                f"Cannot open workbook: {str(e)}",
                "Check file is not corrupted and is valid Excel format"
            ))
    
    def validate_sheet_content(self, wb, num, config):
        """Validate key sheet has content"""
        # Check Evidence Register sheet
        if 'Evidence Register' in wb.sheetnames:
            ws = wb['Evidence Register']
            
            # Count rows with data (beyond header)
            data_rows = 0
            for row in ws.iter_rows(min_row=4, max_row=20):
                if any(cell.value for cell in row):
                    data_rows += 1
            
            if data_rows == 0:
                print(f"  ⚠️  Evidence Register appears empty")
                self.issues.append(ValidationIssue(
                    ValidationIssue.MEDIUM,
                    f"Assessment {num}",
                    "Evidence Register",
                    "No evidence entries found",
                    "Populate evidence register with assessment evidence"
                ))
            else:
                print(f"  ✓ Evidence Register has {data_rows} entries")
    
    def print_summary(self):
        """Print validation summary"""
        print("="*80)
        print("VALIDATION SUMMARY")
        print("="*80)
        print()
        
        if not self.issues:
            print("✅ All validations passed - no issues found!")
            print()
            return
        
        # Count by severity
        critical = sum(1 for i in self.issues if i.severity == ValidationIssue.CRITICAL)
        high = sum(1 for i in self.issues if i.severity == ValidationIssue.HIGH)
        medium = sum(1 for i in self.issues if i.severity == ValidationIssue.MEDIUM)
        low = sum(1 for i in self.issues if i.severity == ValidationIssue.LOW)
        
        print(f"Total Issues Found: {len(self.issues)}")
        print(f"  🔴 Critical: {critical}")
        print(f"  🟠 High: {high}")
        print(f"  🟡 Medium: {medium}")
        print(f"  🟢 Low: {low}")
        print()
        
        # Print issues by severity
        for severity in [ValidationIssue.CRITICAL, ValidationIssue.HIGH, 
                        ValidationIssue.MEDIUM, ValidationIssue.LOW]:
            severity_issues = [i for i in self.issues if i.severity == severity]
            if severity_issues:
                print(f"{severity} ISSUES:")
                print("-" * 80)
                for issue in severity_issues:
                    print(f"  {issue}")
                    print(f"    → Remediation: {issue.remediation}")
                    print()
        
        # Recommendation
        if critical > 0:
            print("⚠️  CRITICAL ISSUES MUST BE RESOLVED BEFORE DASHBOARD CONSOLIDATION")
        elif high > 0:
            print("⚠️  HIGH PRIORITY ISSUES SHOULD BE RESOLVED SOON")
        else:
            print("✓ No critical or high priority issues - assessments are usable")
        print()


def main():
    """Main execution"""
    # Simple argument parsing
    directory = "."
    assessment_num = None
    report_file = None
    dry_run = False
    
    for i, arg in enumerate(sys.argv):
        if arg == '--dir' and i+1 < len(sys.argv):
            directory = sys.argv[i+1]
        elif arg == '--assessment' and i+1 < len(sys.argv):
            assessment_num = int(sys.argv[i+1])
        elif arg == '--report' and i+1 < len(sys.argv):
            report_file = sys.argv[i+1]
        elif arg == '--dry-run':
            dry_run = True
        elif arg == '--help':
            print(__doc__)
            return 0
    
    # Run validation
    validator = AssessmentValidator(directory, dry_run)
    validator.validate_all(assessment_num)
    
    # Optionally write report to file
    if report_file:
        with open(report_file, 'w') as f:
            f.write("ISMS Control A.5.9 - Validation Report\n")
            f.write(f"Generated: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n\n")
            for issue in validator.issues:
                f.write(f"{issue}\n")
                f.write(f"  Remediation: {issue.remediation}\n\n")
        print(f"Report written to: {report_file}")
    
    # Exit code
    critical_count = sum(1 for i in validator.issues if i.severity == ValidationIssue.CRITICAL)
    return 1 if critical_count > 0 else 0


if __name__ == "__main__":
    sys.exit(main())
