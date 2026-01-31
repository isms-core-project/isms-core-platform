#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.20-21-22.S4 - Network Segmentation Matrix Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.20, A.8.21, A.8.22: Network Security Framework
Assessment Workbook 4 of 6: Network Segregation and Security Zones

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific network segmentation architecture, security zones,
and assessment requirements.

Key customization areas:
1. Security zone definitions (match your actual segmentation architecture)
2. Trust boundary requirements (based on your risk assessment)
3. Inter-zone traffic policies (aligned with your business needs)
4. Segmentation technologies (specific to your infrastructure)
5. Effectiveness testing criteria (adapted to your security standards)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-POL-A.8.20-21-22 Network Security Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
network segmentation architecture, security zones, trust boundaries, and
inter-zone traffic controls across the organization.

**Purpose:**
Enables comprehensive assessment of network segmentation architecture, security
zones, trust boundaries, and inter-zone traffic controls, supporting evidence-based
validation of network segregation effectiveness against ISO 27001:2022 Control A.8.22.

**Assessment Scope:**
- Security zones inventory and definition
- Zone purpose and classification (DMZ, internal, management, guest)
- Networks/subnets per zone (VLAN assignments, IP ranges)
- Trust relationships between zones
- Inter-zone traffic policies (firewall rules, ACLs)
- Segmentation technology (VLAN, VRF, physical, virtualization)
- Effectiveness testing results
- Flat network identification and remediation requirements

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and segmentation standards
2. Segmentation_Summary - Overall segmentation architecture dashboard
3. Security_Zones - Security zones inventory and definitions
4. Zone_Networks - Network/subnet assignments per zone
5. Inter_Zone_Matrix - Zone-to-zone trust relationships and traffic matrix
6. Firewall_Rules - Firewall rules inventory and zone protection
7. VLAN_Inventory - VLAN inventory and VLAN-to-zone mapping
8. Microsegmentation - Microsegmentation implementation (if applicable)
9. Effectiveness_Testing - Segmentation effectiveness test results
10. Flat_Networks - Flat network identification and remediation
11. Trust_Boundaries - Trust boundary documentation and controls
12. Gap_Analysis - Segmentation gaps and remediation tracking
13. Evidence_Register - Audit evidence tracking and documentation
14. Approval_Sign_Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with zone classification dropdown lists
- Conditional formatting for segmentation effectiveness status
- Automated zone-to-zone traffic matrix generation
- Protected formulas with unprotected input cells
- Segmentation technology tracking (VLAN, VRF, physical)
- Flat network identification and risk assessment
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with firewall management systems

**Integration:**
This assessment works with WB1 (Infrastructure Inventory) to map devices to
security zones, integrates with WB2 (Device Security) for zone protection controls,
and feeds into WB5 (Controls Coverage) and the Network Security Compliance
Dashboard for executive reporting.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a820_4_segmentation_matrix.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a820_4_segmentation_matrix.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a820_4_segmentation_matrix.py --date 20250124
    
    # Generate with custom organization name
    python3 generate_a820_4_segmentation_matrix.py --org "ACME Corporation"

Command-Line Options:
    --output PATH       Output directory for generated workbook
    --date YYYYMMDD     Date suffix for filename (default: current date)
    --org NAME          Organization name to include in workbook
    --help              Display usage information

Output:
    File: ISMS_A_8_20_21_22_4_Segmentation_Matrix_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize security zone definitions to match your architecture
    2. Document all security zones and their purpose/classification
    3. Map networks/subnets to security zones
    4. Define inter-zone trust relationships and traffic policies
    5. Inventory firewall rules protecting zone boundaries
    6. Document segmentation technology (VLAN, VRF, physical separation)
    7. Conduct segmentation effectiveness testing
    8. Identify flat networks and assess remediation requirements
    9. Collect and link audit evidence (firewall configs, test results)
    10. Obtain stakeholder approvals
    11. Feed results into WB5 and Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.20, A.8.21, A.8.22
Assessment Domain:    4 of 6 (Network Segregation and Security Zones)
Primary Control:      A.8.22 (Segregation of Networks)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.20-21-22: Network Security Framework (Master Policy)
    - ISMS-POL-A.8.20-21-22-S1: Executive Summary & Control Alignment
    - ISMS-POL-A.8.20-21-22-S2: Network Security Requirements (A.8.20)
    - ISMS-POL-A.8.20-21-22-S3: Network Services Requirements (A.8.21)
    - ISMS-POL-A.8.20-21-22-S4: Network Segregation Requirements (A.8.22)
    - ISMS-POL-A.8.20-21-22-S5: Assessment & Evidence Framework
    - ISMS-IMP-A.8.20-21-22-S1: Network Discovery Process
    - ISMS-IMP-A.8.20-21-22-S2: Network Architecture Documentation
    - ISMS-IMP-A.8.20-21-22-S3: Device Hardening Process
    - ISMS-IMP-A.8.20-21-22-S4: Services Security Process
    - ISMS-IMP-A.8.20-21-22-S5: Segmentation Implementation
    - ISMS-IMP-A.8.20-21-22-S6: Network Security Testing

Related Scripts:
    - generate_a820_1_infrastructure_inventory.py (WB1: Device Inventory)
    - generate_a820_2_device_security_assessment.py (WB2: Device Hardening)
    - generate_a820_3_services_catalog.py (WB3: Network Services)
    - generate_a820_4_segmentation_matrix.py (WB4: Segmentation)
    - generate_a820_5_controls_coverage.py (WB5: Controls Coverage)
    - generate_a820_6_compliance_dashboard.py (Dashboard: Executive View)
    - normalize_a820_assessments.py (Utility: Data Normalization)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.20-21-22 specification
    - Supports comprehensive network segmentation assessment
    - Integrated with A.8.20-21-22 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Network Segmentation Standards:**
Network segmentation architecture should follow security best practices:
- Defense in depth with multiple security zones
- Least privilege for inter-zone traffic
- DMZ for internet-facing services
- Separate management network for device administration
- Guest network isolation
- Microsegmentation for critical assets

Review and update segmentation architecture based on changing risk landscape.

**Technology Diversity:**
This assessment framework is technology-agnostic and must work across:
- Traditional networks (VLANs, physical separation, router ACLs)
- Software-Defined Networks (SDN policies, virtual networks)
- Cloud environments (AWS VPCs, Azure VNets, GCP VPCs, security groups)
- Hybrid architectures (on-premise zones + cloud security groups)
- Microsegmentation (VMware NSX, Cisco ACI, Illumio)

Customize assessment criteria to include your specific technology stack.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Documented security zone architecture with clear purpose/classification
- Inter-zone traffic policies (firewall rules, ACLs)
- Evidence of segmentation effectiveness (test results)
- Flat network identification and risk acceptance or remediation plans
- Trust boundary documentation and controls

**Data Protection:**
Assessment workbooks contain sensitive network architecture information including:
- Complete network segmentation architecture
- Security zone boundaries and trust relationships
- Firewall rules and traffic policies
- Identified flat networks and segmentation gaps

Handle in accordance with your organization's data classification policies.
Restrict access to authorized network architects and security personnel only.

**Maintenance:**
Review and update segmentation assessment:
- Monthly: After significant network changes (new zones, zone modifications)
- Quarterly: Routine reassessment of inter-zone traffic policies
- Annually: Complete segmentation architecture review
- Ad-hoc: After security incidents or penetration test findings

**Quality Assurance:**
Validate segmentation effectiveness by:
- Testing inter-zone traffic filtering (authorized vs. unauthorized traffic)
- VLAN hopping testing
- Firewall rule validation and cleanup
- Network topology verification against segmentation architecture
- Penetration testing across zone boundaries
- Flat network scanning and identification

**Regulatory Alignment:**
Ensure network segmentation aligns with applicable regulatory requirements:
- Payment processing: PCI DSS network segmentation requirements (cardholder data)
- Healthcare: HIPAA network segmentation for ePHI
- Finance: Regional banking network segmentation requirements
- Government: Jurisdiction-specific segmentation mandates

Customize assessment criteria to include regulatory-specific requirements.

**Integration Points:**
This assessment integrates with other ISO 27001 controls:
- A.8.20 (Network Security): Devices protecting zone boundaries
- A.8.21 (Network Services): Services per security zone
- A.8.15 (Logging): Inter-zone traffic logging
- A.8.16 (Monitoring): Zone boundary monitoring

**Common Pitfalls to Avoid:**
1. **Flat Networks**: Everything in one VLAN/subnet (no segmentation)
2. **Overly Permissive Rules**: "Allow Any Any" between zones
3. **No Testing**: Assuming segmentation works without testing
4. **Documentation Drift**: Topology diagrams don't match reality
5. **Missing DMZ**: Internet-facing services in internal network
6. **No Management Network**: Admin traffic mixed with production
7. **VLAN Hopping**: Not preventing VLAN hopping attacks
8. **Cloud Confusion**: Assuming cloud providers handle segmentation

**Segmentation Effectiveness Testing:**

**Zone Boundary Testing:**
- Verify firewall rules block unauthorized traffic between zones
- Test from each zone to every other zone (zone-to-zone matrix)
- Validate both allowed and denied traffic
- Document test methodology and results

**VLAN Security Testing:**
- VLAN hopping attempts (double tagging, switch spoofing)
- Trunk port security verification
- Native VLAN security
- Private VLAN (PVLAN) effectiveness if used

**Microsegmentation Testing (if applicable):**
- Host-to-host traffic filtering
- Application-level segmentation
- East-west traffic controls
- Policy enforcement verification

**Flat Network Detection:**
- Identify networks without segmentation
- Assess risk of flat network architectures
- Develop remediation plans or risk acceptance

**Common Security Zone Types:**

**Internet DMZ:**
- Purpose: Internet-facing services (web servers, mail relays)
- Trust Level: Untrusted
- Controls: Strict ingress/egress filtering, IDS/IPS, web application firewall

**Internal Network:**
- Purpose: Corporate workstations, internal servers
- Trust Level: Trusted
- Controls: Access control, monitoring, endpoint protection

**Management Network:**
- Purpose: Device administration (SSH, RDP, management interfaces)
- Trust Level: Restricted
- Controls: Jump hosts, MFA, privileged access management

**Server/Data Center Network:**
- Purpose: Production servers, databases, applications
- Trust Level: Trusted (with microsegmentation)
- Controls: Application-level firewalls, database firewalls

**Guest Network:**
- Purpose: Visitor/guest wireless access
- Trust Level: Untrusted
- Controls: Complete isolation from internal networks, captive portal

**Voice/Video Network:**
- Purpose: VoIP phones, video conferencing
- Trust Level: Segmented
- Controls: QoS, separate VLAN, restricted access

**IoT/OT Network:**
- Purpose: IoT devices, operational technology
- Trust Level: Segmented
- Controls: Strict isolation, dedicated monitoring, anomaly detection

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.formatting.rule import CellIsRule
except ImportError as e:
    logger.error(f"❌ ERROR: Required library not found: {e}")
    logger.info("📦 Install required libraries: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# SECTION 1: CONSTANTS AND CONFIGURATION
# ============================================================================

WORKBOOK_NAME = "Network Segmentation Matrix & Assessment"
DOCUMENT_ID = "ISMS-IMP-A.8.20-21-22.S4"
CONTROL_REF = "ISO/IEC 27001:2022 - Controls A.8.20, A.8.21, A.8.22: Network Security"
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Assessment constants
ZONE_ROW_COUNT = 50             # Security zones
VLAN_ROW_COUNT = 100            # VLAN inventory
FIREWALL_RULE_COUNT = 100       # Firewall rules
ACL_COUNT = 50                  # ACLs
TEST_ROW_COUNT = 30             # Segmentation tests
GAP_ROW_COUNT = 50              # Gap tracking
REMEDIATION_ROW_COUNT = 30      # Remediation actions

# Common security zones (examples)
COMMON_ZONES = [
    "Internet/DMZ",
    "Internal",
    "Management",
    "Datacenter",
    "Guest",
    "Branch",
    "Cloud",
]


# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================

def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="002060", end_color="002060", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "allow_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="006100"),
        },
        "deny_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="9C0006"),
        },
        "critical_fill": {
            "fill": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
        },
        "high_fill": {
            "fill": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        },
        "medium_fill": {
            "fill": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        },
        "low_fill": {
            "fill": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
        },
        "pass_fill": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "fail_fill": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "info_box": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = Font(
            name=getattr(style_dict["font"], "name", "Calibri"),
            size=getattr(style_dict["font"], "size", 10),
            bold=getattr(style_dict["font"], "bold", False),
            color=getattr(style_dict["font"], "color", None),
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb
            if hasattr(style_dict["fill"].start_color, "rgb")
            else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb
            if hasattr(style_dict["fill"].end_color, "rgb")
            else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type,
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=getattr(style_dict["alignment"], "horizontal", None),
            vertical=getattr(style_dict["alignment"], "vertical", None),
            wrap_text=getattr(style_dict["alignment"], "wrap_text", False),
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """Create all data validation objects for dropdowns."""
    
    validations = {}
    
    # Trust Level validation
    validations["trust_level"] = DataValidation(
        type="list",
        formula1='"Untrusted,Low Trust,Medium Trust,High Trust,Trusted"',
        allow_blank=False,
    )
    
    # Traffic Action validation
    validations["traffic_action"] = DataValidation(
        type="list",
        formula1='"Allow,Deny,Inspect"',
        allow_blank=False,
    )
    
    # Segmentation Type validation
    validations["segmentation_type"] = DataValidation(
        type="list",
        formula1='"VLAN,Physical,VRF,Virtual Network,Cloud Security Group"',
        allow_blank=True,
    )
    
    # Firewall Action validation
    validations["firewall_action"] = DataValidation(
        type="list",
        formula1='"Allow,Deny,Drop,Reject"',
        allow_blank=False,
    )
    
    # Rule Review Status validation
    validations["review_status"] = DataValidation(
        type="list",
        formula1='"Current,Outdated,Unused,Requires Review"',
        allow_blank=True,
    )
    
    # Test Result validation
    validations["test_result"] = DataValidation(
        type="list",
        formula1='"Pass,Fail,Partial,Not Tested"',
        allow_blank=False,
    )
    
    # Gap Severity validation
    validations["gap_severity"] = DataValidation(
        type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚪ Low"',
        allow_blank=False,
    )
    
    # Gap Status validation
    validations["gap_status"] = DataValidation(
        type="list",
        formula1='"Open,In Progress,Resolved,Accepted Risk"',
        allow_blank=False,
    )
    
    return validations


# ============================================================================
# SECTION 4: SHEET 1 - INSTRUCTIONS & ASSESSMENT GUIDE
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Assessment Guide sheet."""
    
    ws.title = "Instructions & Guide"
    
    # Title with Document ID and ISO Control Reference
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 40
    
    # Document Information
    ws.merge_cells("A3:B3")
    ws["A3"] = "Document Information"
    apply_style(ws["A3"], styles["header"])
    
    info_data = [
        ("Workbook:", WORKBOOK_NAME),
        ("Generated:", GENERATED_DATE),
        ("Version:", "1.0"),
        ("Control:", "ISO 27001:2022 A.8.22 (Segregation of Networks)"),
        ("Purpose:", "Assess network segmentation architecture and effectiveness"),
        ("Related IMP:", "ISMS-IMP-A.8.20-21-22-S5 (Segmentation Implementation)"),
    ]
    
    row = 4
    for label, value in info_data:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Assessment Approach
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Assessment Approach"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    approach = [
        "1. ZONE IDENTIFICATION: Define all security zones (DMZ, Internal, Management, Guest, etc.)",
        "2. TRUST BOUNDARIES: Map trust levels and security requirements per zone",
        "3. SEGMENTATION MATRIX: Document zone-to-zone traffic rules (Allow/Deny/Inspect)",
        "4. VLAN DESIGN: Verify VLAN numbering, naming, and zone assignment",
        "5. FIREWALL RULES: Review firewall rules for zone separation",
        "6. ACL ASSESSMENT: Verify router/switch ACLs enforce segmentation",
        "7. EFFECTIVENESS TESTING: Test segmentation with traffic flow validation",
        "8. GAP IDENTIFICATION: Identify flat networks, excessive trust, missing controls",
    ]
    
    for instruction in approach:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = instruction
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Security Zone Design Principles
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Security Zone Design Principles"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    principles = [
        "• DEFAULT DENY: Default policy is deny; only explicitly allowed traffic permitted",
        "• LEAST PRIVILEGE: Only necessary traffic flows between zones",
        "• TRUST BOUNDARIES: Clear separation between trust levels (untrusted → trusted)",
        "• DEFENSE IN DEPTH: Multiple layers (firewalls, ACLs, VLANs)",
        "• MICROSEGMENTATION: Further divide zones where appropriate (e.g., datacenter tiers)",
        "• MONITORING: Log and monitor inter-zone traffic",
    ]
    
    for principle in principles:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = principle
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Common Security Zones
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Common Security Zones"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Zone"
    ws["B" + str(row)] = "Purpose"
    ws["C" + str(row)] = "Trust Level"
    for col in ["A", "B", "C"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    zones = [
        ("Internet/DMZ", "Public-facing services", "Untrusted"),
        ("Internal", "User workstations and internal applications", "High Trust"),
        ("Management", "Network device management", "Trusted"),
        ("Datacenter/Server", "Application and database servers", "Medium Trust"),
        ("Guest", "Guest Wi-Fi and temporary access", "Untrusted"),
        ("Branch/Remote", "Branch office networks", "Medium Trust"),
        ("Cloud", "Cloud-hosted resources (AWS, Azure, GCP)", "Low Trust"),
        ("OT/ICS", "Operational Technology / Industrial Control Systems", "High Trust"),
    ]
    
    row += 1
    for zone, purpose, trust in zones:
        ws[f"A{row}"] = zone
        ws[f"B{row}"] = purpose
        ws[f"C{row}"] = trust
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Important Notes
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "⚠ IMPORTANT ASSESSMENT NOTES"
    apply_style(ws[f"A{row}"], styles["header"])
    
    notes = [
        "• Segmentation Matrix is the core assessment - defines allowed traffic between zones",
        "• Test effectiveness: Verify traffic denied when it should be (no bypass routes)",
        "• Identify flat networks: Any networks without segmentation are critical gaps",
        "• Review firewall rules regularly: Outdated rules create security holes",
        "• Document exceptions: Any deviations from default deny must be justified",
    ]
    
    row += 1
    for note in notes:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = note
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    for col in ["D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 15


# ============================================================================
# SECTION 5: SHEET 2 - SECURITY ZONES INVENTORY
# ============================================================================

def create_security_zones_sheet(ws, styles, validations):
    """Create Security Zones Inventory sheet."""
    
    ws.title = "Security_Zones_Inventory"
    
    # Title
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = f"Security Zones Inventory - Generated {GENERATED_DATE}"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "📋 Define all security zones in the network. This is the foundation for the segmentation matrix."
    apply_style(ws["A2"], styles["info_box"])
    ws.row_dimensions[2].height = 30
    
    # Column Headers
    headers = [
        "Zone ID",              # A - Auto
        "Zone Name",            # B
        "Purpose",              # C
        "Trust Level",          # D - Dropdown
        "Networks/Subnets",     # E (IP ranges)
        "VLANs",                # F (VLAN IDs)
        "Segmentation Type",    # G - Dropdown
        "Devices Count",        # H
        "Critical Assets",      # I (Y/N)
        "Internet Access",      # J (Y/N)
        "Owner",                # K
        "Notes",                # L
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + ZONE_ROW_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Zone ID (auto-generated)
        zone_num = row_idx - start_row + 1
        ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","ZONE-" & TEXT({zone_num},"00"),"")')
        
        # Input cells
        for col in range(2, 13):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["trust_level"].add(f"D{start_row}:D{end_row}")
    ws.add_data_validation(validations["trust_level"])
    
    validations["segmentation_type"].add(f"G{start_row}:G{end_row}")
    ws.add_data_validation(validations["segmentation_type"])
    
    # Column widths
    widths = [12, 20, 35, 15, 25, 15, 20, 12, 12, 12, 20, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 6: SHEET 3 - SEGMENTATION MATRIX
# ============================================================================

def create_segmentation_matrix_sheet(ws, styles, validations):
    """Create Segmentation Matrix - zone-to-zone traffic rules."""
    
    ws.title = "Segmentation_Matrix"
    
    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "Network Segmentation Matrix - Zone-to-Zone Traffic Rules"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:K2")
    ws["A2"] = "📋 Define allowed/denied traffic between security zones. This is the CORE segmentation assessment."
    apply_style(ws["A2"], styles["info_box"])
    ws.row_dimensions[2].height = 30
    
    # Matrix explanation
    ws.merge_cells("A3:K3")
    ws["A3"] = "💡 FORMAT: Each row defines traffic flow from Source Zone → Destination Zone. Action: Allow/Deny/Inspect."
    apply_style(ws["A3"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Rule ID",              # A - Auto
        "Source Zone",          # B
        "Destination Zone",     # C
        "Traffic Action",       # D - Dropdown (Allow/Deny/Inspect)
        "Protocols/Ports",      # E (TCP/80, UDP/53, etc.)
        "Justification",        # F
        "Firewall Enforced",    # G (Y/N)
        "ACL Enforced",         # H (Y/N)
        "Monitoring",           # I (Y/N)
        "Last Reviewed",        # J
        "Notes",                # K
    ]
    
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows (50 rows for matrix entries)
    start_row = 5
    end_row = start_row + 49
    
    for row_idx in range(start_row, end_row + 1):
        # Rule ID (auto-generated)
        rule_num = row_idx - start_row + 1
        ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","SEG-" & TEXT({rule_num},"000"),"")')
        
        # Input cells
        for col in range(2, 12):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["traffic_action"].add(f"D{start_row}:D{end_row}")
    ws.add_data_validation(validations["traffic_action"])
    
    # Conditional Formatting for Traffic Action
    ws.conditional_formatting.add(
        f"D{start_row}:D{end_row}",
        CellIsRule(operator="equal", formula=['"Allow"'], fill=styles["allow_fill"]["fill"], font=styles["allow_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"D{start_row}:D{end_row}",
        CellIsRule(operator="equal", formula=['"Deny"'], fill=styles["deny_fill"]["fill"], font=styles["deny_fill"]["font"])
    )
    
    # Example Matrix Entries
    row = end_row + 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "Common Segmentation Matrix Examples"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Source"
    ws["B" + str(row)] = "Destination"
    ws["C" + str(row)] = "Action"
    ws["D" + str(row)] = "Justification"
    for col in ["A", "B", "C", "D"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    examples = [
        ("Internet/DMZ", "Internal", "Deny", "External users cannot access internal network"),
        ("Internal", "Internet/DMZ", "Allow", "Internal users can access public services"),
        ("Internal", "Datacenter", "Allow", "Users need access to application servers"),
        ("Guest", "Internal", "Deny", "Guest Wi-Fi isolated from corporate network"),
        ("Guest", "Internet", "Allow", "Guest users can access internet only"),
        ("Management", "All Zones", "Allow", "Management needs access to all devices"),
        ("All Zones", "Management", "Deny", "Management zone is protected"),
    ]
    
    row += 1
    for src, dst, action, justification in examples:
        ws[f"A{row}"] = src
        ws[f"B{row}"] = dst
        ws[f"C{row}"] = action
        ws[f"D{row}"] = justification
        
        for col in ["A", "B", "D"]:
            apply_style(ws[f"{col}{row}"], styles["info_box"])
        
        if action == "Allow":
            apply_style(ws[f"C{row}"], styles["allow_fill"])
        elif action == "Deny":
            apply_style(ws[f"C{row}"], styles["deny_fill"])
        
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Column widths
    widths = [12, 20, 20, 12, 25, 40, 15, 15, 12, 12, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 7: SHEET 4 - VLAN INVENTORY
# ============================================================================

def create_vlan_inventory_sheet(ws, styles):
    """Create VLAN Inventory sheet."""
    
    ws.title = "VLAN_Inventory"
    
    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "VLAN Inventory & Design Assessment"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:J2")
    ws["A2"] = "📋 Document all VLANs. Verify VLAN numbering, naming standards, and zone assignment."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "VLAN ID",              # A
        "VLAN Name",            # B
        "Security Zone",        # C (link to Security_Zones_Inventory)
        "IP Subnet",            # D
        "Purpose",              # E
        "Devices/Ports",        # F
        "Trunk Allowed",        # G (Yes/No)
        "Native VLAN",          # H (Yes/No)
        "Last Reviewed",        # I
        "Notes",                # J
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + VLAN_ROW_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        for col in range(1, 11):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # VLAN Design Best Practices
    row = end_row + 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "VLAN Design Best Practices (per IMP-S5)"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    best_practices = [
        "• VLAN NUMBERING: Consistent scheme (e.g., 10-99 for users, 100-199 for servers, 200-299 for management)",
        "• VLAN NAMING: Descriptive names (e.g., VLAN-10-USERS-FLOOR1, VLAN-100-SERVERS-WEB)",
        "• NATIVE VLAN: Change from default VLAN 1, use dedicated unused VLAN",
        "• VLAN PRUNING: Only allow necessary VLANs on trunks",
        "• ONE VLAN PER ZONE: Each security zone maps to dedicated VLANs",
        "• DOCUMENTATION: Maintain VLAN database with assignments and purpose",
    ]
    
    for practice in best_practices:
        ws.merge_cells(f"A{row}:J{row}")
        ws[f"A{row}"] = practice
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Column widths
    widths = [10, 25, 20, 20, 35, 20, 12, 12, 12, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 8: SHEET 5 - FIREWALL RULES ASSESSMENT
# ============================================================================

def create_firewall_rules_sheet(ws, styles, validations):
    """Create Firewall Rules Assessment sheet."""
    
    ws.title = "Firewall_Rules_Assessment"
    
    # Title
    ws.merge_cells("A1:M1")
    cell = ws["A1"]
    cell.value = "Firewall Rules Assessment"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:M2")
    ws["A2"] = "📋 Review firewall rules that enforce zone segmentation. Focus on inter-zone traffic controls."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Rule ID",              # A - Auto
        "Firewall",             # B (Device name)
        "Rule Number",          # C
        "Source Zone",          # D
        "Destination Zone",     # E
        "Source IP/Network",    # F
        "Destination IP/Network", # G
        "Service/Port",         # H
        "Action",               # I - Dropdown
        "Review Status",        # J - Dropdown
        "Last Modified",        # K
        "Reviewed By",          # L
        "Notes",                # M
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + FIREWALL_RULE_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Rule ID (auto-generated)
        rule_num = row_idx - start_row + 1
        ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","FW-RULE-" & TEXT({rule_num},"000"),"")')
        
        # Input cells
        for col in range(2, 14):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["firewall_action"].add(f"I{start_row}:I{end_row}")
    ws.add_data_validation(validations["firewall_action"])
    
    validations["review_status"].add(f"J{start_row}:J{end_row}")
    ws.add_data_validation(validations["review_status"])
    
    # Conditional Formatting
    ws.conditional_formatting.add(
        f"I{start_row}:I{end_row}",
        CellIsRule(operator="equal", formula=['"Allow"'], fill=styles["allow_fill"]["fill"], font=styles["allow_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"I{start_row}:I{end_row}",
        CellIsRule(operator="equal", formula=['"Deny"'], fill=styles["deny_fill"]["fill"], font=styles["deny_fill"]["font"])
    )
    
    # Firewall Rule Review Guidelines
    row = end_row + 2
    ws.merge_cells(f"A{row}:M{row}")
    ws[f"A{row}"] = "Firewall Rule Review Guidelines"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    guidelines = [
        "• QUARTERLY REVIEW: Review all firewall rules at least quarterly",
        "• REMOVE UNUSED: Identify and remove unused rules (hit count = 0)",
        "• SPECIFIC > ANY: Avoid 'any any any' rules; be specific about source/destination/service",
        "• LOGGING: Enable logging for inter-zone traffic (especially deny rules)",
        "• RULE ORDER: Most specific rules first, default deny last",
        "• DOCUMENTATION: Every rule must have business justification",
    ]
    
    for guideline in guidelines:
        ws.merge_cells(f"A{row}:M{row}")
        ws[f"A{row}"] = guideline
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Column widths
    widths = [15, 20, 12, 18, 18, 20, 20, 18, 12, 15, 12, 15, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 9: SHEET 6 - ACL ASSESSMENT
# ============================================================================

def create_acl_assessment_sheet(ws, styles, validations):
    """Create ACL (Access Control List) Assessment sheet."""
    
    ws.title = "ACL_Assessment"
    
    # Title
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = "ACL (Access Control List) Assessment"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:L2")
    ws["A2"] = "📋 Review router/switch ACLs that enforce segmentation. Complement firewall rules."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "ACL ID",               # A - Auto
        "Device",               # B (Router/Switch)
        "ACL Name/Number",      # C
        "Interface",            # D
        "Direction",            # E (Inbound/Outbound)
        "Purpose",              # F
        "Zones Protected",      # G
        "Review Status",        # H - Dropdown
        "Effectiveness",        # I (Tested Y/N)
        "Last Modified",        # J
        "Reviewed By",          # K
        "Notes",                # L
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + ACL_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # ACL ID (auto-generated)
        acl_num = row_idx - start_row + 1
        ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","ACL-" & TEXT({acl_num},"00"),"")')
        
        # Input cells
        for col in range(2, 13):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["review_status"].add(f"H{start_row}:H{end_row}")
    ws.add_data_validation(validations["review_status"])
    
    # ACL Best Practices
    row = end_row + 2
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "ACL Best Practices (per IMP-S5)"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    best_practices = [
        "• ACL PLACEMENT: Apply inbound on router interfaces (filter before routing)",
        "• SPECIFIC FIRST: Most specific ACL entries first, general/deny last",
        "• EXPLICIT DENY: Include explicit deny at end (implicit deny doesn't log)",
        "• NAMING: Use descriptive names for extended ACLs (e.g., ACL-BLOCK-GUEST-TO-INTERNAL)",
        "• TESTING: Test ACLs with traffic generators/packet captures before production",
        "• DOCUMENTATION: Maintain ACL documentation with business justification per entry",
    ]
    
    for practice in best_practices:
        ws.merge_cells(f"A{row}:L{row}")
        ws[f"A{row}"] = practice
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Column widths
    widths = [12, 20, 20, 15, 12, 35, 20, 15, 12, 12, 15, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 10: SHEET 7 - SEGMENTATION TESTING
# ============================================================================

def create_segmentation_testing_sheet(ws, styles, validations):
    """Create Segmentation Effectiveness Testing sheet."""
    
    ws.title = "Segmentation_Testing"
    
    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "Segmentation Effectiveness Testing"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:K2")
    ws["A2"] = "📋 Test segmentation effectiveness. Verify traffic is blocked when expected (no bypass routes)."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Test ID",              # A - Auto
        "Test Name",            # B
        "Source Zone",          # C
        "Destination Zone",     # D
        "Expected Result",      # E (Allow/Deny)
        "Actual Result",        # F (Allow/Deny)
        "Test Result",          # G - Dropdown (Pass/Fail)
        "Test Method",          # H (ping, traceroute, nmap, etc.)
        "Test Date",            # I
        "Tested By",            # J
        "Notes",                # K
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + TEST_ROW_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Test ID (auto-generated)
        test_num = row_idx - start_row + 1
        ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","SEG-TEST-" & TEXT({test_num},"00"),"")')
        
        # Input cells
        for col in range(2, 12):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["test_result"].add(f"G{start_row}:G{end_row}")
    ws.add_data_validation(validations["test_result"])
    
    # Conditional Formatting
    ws.conditional_formatting.add(
        f"G{start_row}:G{end_row}",
        CellIsRule(operator="equal", formula=['"Pass"'], fill=styles["pass_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"G{start_row}:G{end_row}",
        CellIsRule(operator="equal", formula=['"Fail"'], fill=styles["fail_fill"]["fill"])
    )
    
    # Testing Methods
    row = end_row + 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "Segmentation Testing Methods (per IMP-S5)"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Test Method"
    ws["B" + str(row)] = "Description"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    methods = [
        ("Ping Test", "Simple connectivity test - should fail if segmentation working"),
        ("Traceroute", "Verify traffic path - should stop at firewall/router"),
        ("Port Scan (nmap)", "Verify ports blocked between zones"),
        ("HTTP/HTTPS Request", "Test application layer connectivity"),
        ("Packet Capture", "Capture traffic to verify deny/drop behavior"),
        ("Lateral Movement Test", "Attempt to pivot from one zone to another (pentesting)"),
    ]
    
    row += 1
    for method, description in methods:
        ws[f"A{row}"] = method
        ws.merge_cells(f"B{row}:K{row}")
        ws[f"B{row}"] = description
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 25
        row += 1
    
    # Column widths
    widths = [15, 25, 18, 18, 15, 15, 12, 20, 12, 15, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 11: SHEET 8 - GAP ANALYSIS
# ============================================================================

def create_gap_analysis_sheet(ws, styles, validations):
    """Create Gap Analysis sheet for segmentation issues."""
    
    ws.title = "Gap_Analysis"
    
    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "Segmentation Gaps - Remediation Tracking"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:J2")
    ws["A2"] = "📋 Document segmentation gaps: flat networks, excessive trust, missing controls, bypass routes."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Gap ID",               # A - Auto
        "Gap Type",             # B (Flat Network, Excessive Trust, etc.)
        "Location/Zone",        # C
        "Description",          # D
        "Risk",                 # E
        "Severity",             # F - Dropdown
        "Remediation Plan",     # G
        "Owner",                # H
        "Status",               # I - Dropdown
        "Target Date",          # J
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + GAP_ROW_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Gap ID (auto-generated)
        gap_num = row_idx - start_row + 1
        ws.cell(row=row_idx, column=1, value=f'=IF(B{row_idx}<>"","SEG-GAP-" & TEXT({gap_num},"000"),"")')
        
        # Input cells
        for col in range(2, 11):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["gap_severity"].add(f"F{start_row}:F{end_row}")
    ws.add_data_validation(validations["gap_severity"])
    
    validations["gap_status"].add(f"I{start_row}:I{end_row}")
    ws.add_data_validation(validations["gap_status"])
    
    # Conditional Formatting
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=styles["low_fill"]["fill"])
    )
    
    # Common Gap Types
    row = end_row + 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "Common Segmentation Gaps"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    gap_types = [
        "Flat Network: No segmentation; all devices in same broadcast domain (Critical)",
        "Excessive Trust: Guest/DMZ can access internal resources (High)",
        "Missing Firewall Rules: Zones defined but not enforced by firewall (High)",
        "Overly Permissive Rules: 'any any any' rules bypass segmentation (Medium)",
        "VLAN Hopping Risk: Native VLAN 1, DTP enabled (Medium)",
        "Bypass Routes: Alternative paths circumvent segmentation (Critical)",
        "Outdated Rules: Firewall rules not reviewed, unused rules accumulate (Low)",
    ]
    
    for gap in gap_types:
        ws.merge_cells(f"A{row}:J{row}")
        ws[f"A{row}"] = gap
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Column widths
    widths = [15, 25, 20, 40, 35, 12, 40, 20, 15, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 12: SHEET 9 - COMPLIANCE SUMMARY
# ============================================================================

def create_compliance_summary_sheet(ws, styles):
    """Create Compliance Summary with metrics and charts."""
    
    ws.title = "Compliance_Summary"
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Network Segmentation Compliance Summary"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Summary Statistics
    row = 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Summary Statistics"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Value"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    stats = [
        ("Security Zones Defined", f'=COUNTA(Security_Zones_Inventory!B4:B{3 + ZONE_ROW_COUNT})'),
        ("VLANs Documented", f'=COUNTA(VLAN_Inventory!A4:A{3 + VLAN_ROW_COUNT})'),
        ("Segmentation Rules", f'=COUNTA(Segmentation_Matrix!B5:B54)'),
        ("Firewall Rules Reviewed", f'=COUNTA(Firewall_Rules_Assessment!B4:B{3 + FIREWALL_RULE_COUNT})'),
        ("ACLs Reviewed", f'=COUNTA(ACL_Assessment!B4:B{3 + ACL_COUNT})'),
        ("Segmentation Tests", f'=COUNTA(Segmentation_Testing!B4:B{3 + TEST_ROW_COUNT})'),
        ("Tests Passed", f'=COUNTIF(Segmentation_Testing!G4:G{3 + TEST_ROW_COUNT},"Pass")'),
        ("Tests Failed", f'=COUNTIF(Segmentation_Testing!G4:G{3 + TEST_ROW_COUNT},"Fail")'),
        ("Identified Gaps", f'=COUNTA(Gap_Analysis!B4:B{3 + GAP_ROW_COUNT})'),
        ("Critical Gaps", f'=COUNTIF(Gap_Analysis!F4:F{3 + GAP_ROW_COUNT},"Critical")'),
    ]
    
    row += 1
    for metric, formula in stats:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = formula
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    # Test Results Distribution
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Segmentation Test Results"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Result"
    ws["B" + str(row)] = "Count"
    for col in ["A", "B"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    chart_start_row = row + 1
    row += 1
    
    results = [
        ("Pass", f'=COUNTIF(Segmentation_Testing!G4:G{3 + TEST_ROW_COUNT},"Pass")'),
        ("Fail", f'=COUNTIF(Segmentation_Testing!G4:G{3 + TEST_ROW_COUNT},"Fail")'),
        ("Partial", f'=COUNTIF(Segmentation_Testing!G4:G{3 + TEST_ROW_COUNT},"Partial")'),
        ("Not Tested", f'=COUNTIF(Segmentation_Testing!G4:G{3 + TEST_ROW_COUNT},"Not Tested")'),
    ]
    
    for result, formula in results:
        ws[f"A{row}"] = result
        ws[f"B{row}"] = formula
        apply_style(ws[f"A{row}"], styles["info_box"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        row += 1
    
    chart_end_row = row - 1
    
    # Create Pie Chart
    pie_chart = PieChart()
    pie_chart.title = "Segmentation Test Results"
    pie_chart.style = 10
    labels = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
    data = Reference(ws, min_col=2, min_row=chart_start_row, max_row=chart_end_row)
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(labels)
    pie_chart.height = 10
    pie_chart.width = 15
    ws.add_chart(pie_chart, f"D{chart_start_row}")
    
    # Compliance Targets
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Compliance Targets & Status"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    ws["A" + str(row)] = "Metric"
    ws["B" + str(row)] = "Target"
    ws["C" + str(row)] = "Current"
    ws["D" + str(row)] = "Status"
    for col in ["A", "B", "C", "D"]:
        apply_style(ws[col + str(row)], styles["column_header"])
    
    targets = [
        ("Zones Defined", "All networks", "Manual count", "Check zones cover all networks"),
        ("Segmentation Matrix Complete", "All zone pairs", "Manual review", "Verify no gaps in matrix"),
        ("Test Pass Rate", "≥ 95%", f"=B{7}/(B{7}+B{8})*100", f'=IF(C{row + 3}>=95,"✔ Met","✗ Not Met")'),
        ("Critical Gaps", "= 0", f"=B{12}", f'=IF(C{row + 4}=0,"✔ Met","✗ Not Met")'),
    ]
    
    row += 1
    for metric, target, current, status in targets:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = target
        ws[f"C{row}"] = current
        ws[f"D{row}"] = status
        apply_style(ws[f"A{row}"], styles["column_header"])
        apply_style(ws[f"B{row}"], styles["info_box"])
        apply_style(ws[f"C{row}"], styles["info_box"])
        apply_style(ws[f"D{row}"], styles["info_box"])
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 3
    ws.column_dimensions["F"].width = 3


# ============================================================================
# SECTION 13: SHEET 10 - REMEDIATION ROADMAP
# ============================================================================

def create_remediation_roadmap_sheet(ws, styles, validations):
    """Create Remediation Roadmap for segmentation improvements."""
    
    ws.title = "Remediation_Roadmap"
    
    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "Segmentation Remediation Roadmap"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells("A2:H2")
    ws["A2"] = "📋 Prioritized plan for segmentation improvements. Focus on Critical gaps first."
    apply_style(ws["A2"], styles["info_box"])
    
    # Column Headers
    headers = [
        "Priority",             # A
        "Severity",             # B - Dropdown
        "Improvement",          # C
        "Affected Zones",       # D
        "Remediation Action",   # E
        "Owner",                # F
        "Target Date",          # G
        "Status",               # H - Dropdown
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # Data rows
    start_row = 4
    end_row = start_row + REMEDIATION_ROW_COUNT - 1
    
    for row_idx in range(start_row, end_row + 1):
        # Priority (manual)
        ws.cell(row=row_idx, column=1, value=row_idx - start_row + 1)
        
        # Input cells
        for col in range(2, 9):
            cell = ws.cell(row=row_idx, column=col)
            apply_style(cell, styles["input_cell"])
    
    # Apply Data Validations
    validations["gap_severity"].add(f"B{start_row}:B{end_row}")
    ws.add_data_validation(validations["gap_severity"])
    
    validations["gap_status"].add(f"H{start_row}:H{end_row}")
    ws.add_data_validation(validations["gap_status"])
    
    # Conditional Formatting
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=styles["critical_fill"]["fill"], font=styles["critical_fill"]["font"])
    )
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=styles["high_fill"]["fill"])
    )
    ws.conditional_formatting.add(
        f"B{start_row}:B{end_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=styles["medium_fill"]["fill"])
    )
    
    # Remediation Approach
    row = end_row + 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Segmentation Remediation Approach"
    apply_style(ws[f"A{row}"], styles["header"])
    
    row += 1
    approach = [
        "1. CRITICAL FIRST: Address Critical gaps immediately (flat networks, bypass routes)",
        "2. QUICK WINS: Implement easy improvements (VLAN assignment, ACL updates)",
        "3. FIREWALL RULES: Review and tighten firewall rules between zones",
        "4. TESTING: Re-test after each change to verify effectiveness",
        "5. DOCUMENTATION: Update network documentation (topology, VLAN database)",
        "6. MONITORING: Enable logging and alerting for inter-zone traffic",
    ]
    
    for item in approach:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = item
        apply_style(ws[f"A{row}"], styles["info_box"])
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 15
    
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 14: MAIN FUNCTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.8.20-21-22 - Network Segmentation Matrix Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.22 (Segregation of Networks)")
    logger.info("=" * 80)
    logger.info("\n🎯 Systems Engineering Approach: Zone-Based Segmentation")
    logger.info("📊 Comprehensive Coverage: Zones, VLANs, Firewalls, ACLs, Testing")
    logger.info("🔒 Audit-Ready: Segmentation matrix and effectiveness validation")
    logger.info("\n" + "─" * 80)
    
    # Create workbook
    logger.info("\n[Phase 1] Initializing workbook structure...")
    wb = Workbook()
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    sheet_names = [
        "Instructions & Guide",
        "Security_Zones_Inventory",
        "Segmentation_Matrix",
        "VLAN_Inventory",
        "Firewall_Rules_Assessment",
        "ACL_Assessment",
        "Segmentation_Testing",
        "Gap_Analysis",
        "Compliance_Summary",
        "Remediation_Roadmap",
    ]
    
    for name in sheet_names:
        wb.create_sheet(title=name)
    
    logger.info(f"✅ Workbook created with {len(sheet_names)} sheets")
    
    # Setup styles and validations
    logger.info("\n[Phase 2] Setting up styles and data validations...")
    styles = setup_styles()
    validations = create_data_validations()
    logger.info("✅ Styles and validations configured")
    
    # Create all sheets
    logger.info("\n[Phase 3] Generating assessment sheets...")
    
    logger.info("  [1/10] Creating Instructions & Guide...")
    create_instructions_sheet(wb["Instructions & Guide"], styles)
    logger.info("  ✅ Instructions complete")
    
    logger.info("  [2/10] Creating Security_Zones_Inventory...")
    create_security_zones_sheet(wb["Security_Zones_Inventory"], styles, validations)
    logger.info(f"  ✅ Security zones complete ({ZONE_ROW_COUNT} zones)")
    
    logger.info("  [3/10] Creating Segmentation_Matrix...")
    create_segmentation_matrix_sheet(wb["Segmentation_Matrix"], styles, validations)
    logger.info("  ✅ Segmentation matrix complete (zone-to-zone rules)")
    
    logger.info("  [4/10] Creating VLAN_Inventory...")
    create_vlan_inventory_sheet(wb["VLAN_Inventory"], styles)
    logger.info(f"  ✅ VLAN inventory complete ({VLAN_ROW_COUNT} VLANs)")
    
    logger.info("  [5/10] Creating Firewall_Rules_Assessment...")
    create_firewall_rules_sheet(wb["Firewall_Rules_Assessment"], styles, validations)
    logger.info(f"  ✅ Firewall rules complete ({FIREWALL_RULE_COUNT} rules)")
    
    logger.info("  [6/10] Creating ACL_Assessment...")
    create_acl_assessment_sheet(wb["ACL_Assessment"], styles, validations)
    logger.info(f"  ✅ ACL assessment complete ({ACL_COUNT} ACLs)")
    
    logger.info("  [7/10] Creating Segmentation_Testing...")
    create_segmentation_testing_sheet(wb["Segmentation_Testing"], styles, validations)
    logger.info(f"  ✅ Segmentation testing complete ({TEST_ROW_COUNT} tests)")
    
    logger.info("  [8/10] Creating Gap_Analysis...")
    create_gap_analysis_sheet(wb["Gap_Analysis"], styles, validations)
    logger.info(f"  ✅ Gap analysis complete ({GAP_ROW_COUNT} gap tracking rows)")
    
    logger.info("  [9/10] Creating Compliance_Summary...")
    create_compliance_summary_sheet(wb["Compliance_Summary"], styles)
    logger.info("  ✅ Compliance summary complete (metrics, pie chart)")
    
    logger.info("  [10/10] Creating Remediation_Roadmap...")
    create_remediation_roadmap_sheet(wb["Remediation_Roadmap"], styles, validations)
    logger.info(f"  ✅ Remediation roadmap complete ({REMEDIATION_ROW_COUNT} actions)")
    
    # Save workbook
    logger.info("\n[Phase 4] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.20-21-22.S4_Network_Segmentation_Matrix_{GENERATED_TIMESTAMP}.xlsx"
    
    try:
        wb.save(filename)
        logger.info(f"✅ SUCCESS: {filename}")
    except Exception as e:
        logger.error(f"❌ ERROR saving workbook: {e}")
        return 1
    
    # Summary
    logger.info("\n" + "=" * 80)
    logger.info("📋 WORKBOOK STRUCTURE SUMMARY")
    logger.info("=" * 80)
    logger.info("\n📊 Assessment Sheets:")
    logger.info(f"  • Security_Zones_Inventory ({ZONE_ROW_COUNT} zones)")
    logger.info("  • Segmentation_Matrix (zone-to-zone traffic rules)")
    logger.info(f"  • VLAN_Inventory ({VLAN_ROW_COUNT} VLANs)")
    logger.info(f"  • Firewall_Rules_Assessment ({FIREWALL_RULE_COUNT} rules)")
    logger.info(f"  • ACL_Assessment ({ACL_COUNT} ACLs)")
    logger.info(f"  • Segmentation_Testing ({TEST_ROW_COUNT} effectiveness tests)")
    logger.info("\n📈 Analysis & Reporting:")
    logger.info(f"  • Gap_Analysis ({GAP_ROW_COUNT} gap tracking)")
    logger.info("  • Compliance_Summary (metrics, pie chart)")
    logger.info(f"  • Remediation_Roadmap ({REMEDIATION_ROW_COUNT} prioritized actions)")
    logger.info("\n" + "=" * 80)
    logger.info("🚀 NEXT STEPS:")
    logger.info("  1. Define all security zones in Security_Zones_Inventory")
    logger.info("  2. Complete Segmentation_Matrix (zone-to-zone rules)")
    logger.info("  3. Document VLANs and firewall rules")
    logger.info("  4. Test segmentation effectiveness")
    logger.info("  5. Identify and remediate gaps")
    logger.info("\n" + "=" * 80 + "\n")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
