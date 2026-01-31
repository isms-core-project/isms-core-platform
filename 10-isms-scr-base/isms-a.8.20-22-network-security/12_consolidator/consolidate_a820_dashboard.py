#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
ISMS-A.8.20-22 Network Security - Dashboard Consolidation Utility
Auto-detects workbooks in ../90_workbooks/ and consolidates data into dashboard.
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import os
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Library Imports
# =============================================================================
from openpyxl import load_workbook

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

SOURCE_WORKBOOKS = [
    ('A.8.20-21-22.1', 'Network_Infrastructure', 'Network Infrastructure'),
    ('A.8.20-21-22.2', 'Network_Device_Security', 'Device Security'),
    ('A.8.20-21-22.3', 'Network_Services', 'Network Services'),
    ('A.8.20-21-22.4', 'Network_Segmentation', 'Segmentation'),
    ('A.8.20-21-22.5', 'Network_Controls', 'Controls Coverage'),
]

DASHBOARD_PATTERNS = ['A.8.20-21-22.6', 'A.8.20.6', 'Network_Dashboard', 'Compliance_Dashboard']


def find_workbook(directory, patterns):
    if isinstance(patterns, str):
        patterns = [patterns]
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            for pattern in patterns:
                if pattern in filename:
                    return os.path.join(directory, filename)
    return None


def find_workbooks_by_patterns(directory, pattern_list):
    found = []
    for patterns in pattern_list:
        if isinstance(patterns, tuple):
            wb = find_workbook(directory, [patterns[0], patterns[1]])
            if wb:
                found.append((wb, patterns[2]))
    return found


def safely_write_data(ws, start_row, data):
    count = 0
    for i, row_data in enumerate(data):
        for j, value in enumerate(row_data):
            if value is not None:
                ws.cell(row=start_row + i, column=j + 1, value=value)
        count += 1
    return count


def extract_gaps_from_workbook(filepath, assessment_area, gap_counter_start=1):
    gaps = []
    gap_counter = gap_counter_start
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if any(x in sheet_name.lower() for x in ['gap', 'finding', 'issue']):
                ws = wb[sheet_name]
                for row in range(8, min(ws.max_row + 1, 100)):
                    row_values = [ws.cell(row=row, column=col).value for col in range(1, 12)]
                    if not any(row_values):
                        continue
                    status = gap_desc = None
                    for val in row_values:
                        if val:
                            val_str = str(val)
                            if any(x in val_str for x in ['⚠️', '❌', 'Non-Compliant', 'Partial']):
                                status = val_str
                            elif len(val_str) > 30:
                                gap_desc = val_str
                    if status or gap_desc:
                        severity = 'High' if status and '❌' in str(status) else 'Medium'
                        gaps.append([f'GAP-NET-{gap_counter:03d}', assessment_area,
                            os.path.basename(filepath), row_values[0] or 'Network Device',
                            gap_desc[:200] if gap_desc else f'{assessment_area} gap',
                            status[:30] if status else 'Gap', 'Compliant', severity, 'Open',
                            'Remediation required', None, None])
                        gap_counter += 1
        wb.close()
    except Exception as e:
        logger.warning("Error extracting gaps: %s", e)
    return gaps, gap_counter


def extract_evidence_from_workbook(filepath, assessment_area):
    evidence = []
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if 'evidence' in sheet_name.lower():
                ws = wb[sheet_name]
                for row in range(10, min(ws.max_row + 1, 50)):
                    evd_id = ws.cell(row=row, column=1).value
                    if evd_id and 'EVD' in str(evd_id):
                        evidence.append([evd_id, assessment_area, os.path.basename(filepath), None,
                            ws.cell(row=row, column=2).value, ws.cell(row=row, column=3).value,
                            ws.cell(row=row, column=4).value, ws.cell(row=row, column=5).value,
                            ws.cell(row=row, column=6).value, ws.cell(row=row, column=7).value,
                            ws.cell(row=row, column=8).value, ws.cell(row=row, column=9).value])
        wb.close()
    except Exception as e:
        logger.warning("Error extracting evidence: %s", e)
    return evidence


def generate_risks_and_remediation(gaps):
    risks, remediation = [], []
    today = datetime.now()
    for i, gap in enumerate([g for g in gaps if g[7] == 'High'], 1):
        risks.append([f'RISK-{i:03d}', gap[0], 'Network Security', f'Risk: {gap[4][:80]}',
            gap[3], 'High', 'High', 'High', 'Open', f'Mitigate {gap[0]}', None, None, None])
    for i, gap in enumerate(gaps, 1):
        days = 30 if gap[7] == 'High' else 60
        remediation.append([f'REM-{i:03d}', gap[0], gap[1], gap[3], 'Remediation required',
            'Critical' if gap[7] == 'High' else 'High', 'Planned', None,
            today.strftime('%Y-%m-%d'), (today + timedelta(days=days)).strftime('%Y-%m-%d'),
            None, '0%', None])
    return risks, remediation


def consolidate_dashboard():
    logger.info("=" * 80)
    logger.info("ISMS-A.8.20-22 Network Security - Dashboard Consolidation")
    logger.info("=" * 80)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    workbooks_dir = os.path.join(os.path.dirname(script_dir), '90_workbooks')

    if not os.path.exists(workbooks_dir):
        logger.error("Workbooks directory not found: %s", workbooks_dir)
        return False

    logger.info("Workbooks directory: %s", workbooks_dir)

    dashboard_path = find_workbook(workbooks_dir, DASHBOARD_PATTERNS)
    if not dashboard_path:
        logger.error("Dashboard not found")
        return False
    logger.info("Dashboard: %s", os.path.basename(dashboard_path))

    logger.info("[1/4] Finding source workbooks...")
    sources = find_workbooks_by_patterns(workbooks_dir, SOURCE_WORKBOOKS)
    for path, area in sources:
        logger.info("  Found %s: %s", area, os.path.basename(path))

    logger.info("[2/4] Extracting gaps and evidence...")
    all_gaps, all_evidence = [], []
    gap_counter = 1
    for filepath, area in sources:
        gaps, gap_counter = extract_gaps_from_workbook(filepath, area, gap_counter)
        evidence = extract_evidence_from_workbook(filepath, area)
        logger.info("  %s: %d gaps, %d evidence", area, len(gaps), len(evidence))
        all_gaps.extend(gaps)
        all_evidence.extend(evidence)

    logger.info("[3/4] Generating risks and remediation...")
    all_risks, all_remediation = generate_risks_and_remediation(all_gaps)
    logger.info("  Risks: %d | Remediation: %d", len(all_risks), len(all_remediation))

    logger.info("[4/4] Writing to dashboard...")
    try:
        wb = load_workbook(dashboard_path)
        for sheet_name, data, start_row in [
            ('Gap Analysis', all_gaps, 14), ('Gap_Analysis', all_gaps, 14),
            ('Risk Register', all_risks, 21), ('Remediation Roadmap', all_remediation, 37),
            ('Evidence Register', all_evidence, 14), ('Evidence_Register', all_evidence, 14),
        ]:
            if sheet_name in wb.sheetnames and data:
                count = safely_write_data(wb[sheet_name], start_row, data)
                logger.info("  %s: %d entries", sheet_name, count)
        wb.save(dashboard_path)
        logger.info("Saved: %s", os.path.basename(dashboard_path))
        wb.close()
    except Exception as e:
        logger.error("Error writing to dashboard: %s", e)
        return False

    logger.info("CONSOLIDATION COMPLETE")
    return True


def main():
    """Main entry point with proper error handling."""
    try:
        success = consolidate_dashboard()
        return 0 if success else 1
    except KeyboardInterrupt:
        logger.warning("Operation cancelled by user")
        return 130
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_NOTE: Added license header, logging, import sections, try/except main()
# QA_TOOL: Claude Code Deep Scan
# =============================================================================
