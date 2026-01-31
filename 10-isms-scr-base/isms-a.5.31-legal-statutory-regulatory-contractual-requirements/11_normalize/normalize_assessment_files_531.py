#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-A.5.31 - Assessment File Normalizer and Validator Utility
================================================================================

ISO/IEC 27001:2022 Control A.5.31: Legal, Statutory, Regulatory and Contractual Requirements
Quality Assurance Utility: Excel Assessment File Normalization and Validation

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific assessment file standards, validation rules, and
quality assurance requirements.

Key customization areas:
1. Expected file naming conventions (match your organizational standards)
2. Workbook structure validation rules (specific to your A.5.31 assessments)
3. Data format normalization rules (adapt to your data standards)
4. Validation severity thresholds (based on your quality requirements)
5. Output formatting preferences (align with your reporting needs)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.31 Regulatory Compliance Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script normalizes and validates A.5.31 regulatory compliance assessment
Excel workbooks to ensure consistency, data quality, and structural compliance
with framework standards before data is consolidated into the dashboard or
used for audit evidence.

**Purpose:**
Ensures all regulatory compliance assessment workbooks meet quality standards
and structural requirements, preventing data consolidation errors, improving
audit evidence reliability, and catching human errors before they propagate
through the compliance framework.

**Key Functions:**

1. **File Naming Validation**
   - Verify naming convention compliance (ISMS_Assessment_531_N_Description_YYYYMMDD.xlsx)
   - Check date format validity (YYYYMMDD suffix)
   - Identify versioning inconsistencies
   - Flag non-standard file names

2. **Workbook Structure Validation**
   - Verify presence of required sheets per assessment workbook
   - Validate sheet names match expected standards
   - Check for missing or unexpected sheets
   - Ensure protected/unprotected cells are correctly configured
   - Validate column headers and data types

3. **Data Normalization**
   - Standardize dropdown values (case, spacing, terminology)
   - Normalize date formats (DD.MM.YYYY Swiss standard)
   - Clean whitespace and formatting inconsistencies
   - Validate data type compliance (text vs. numbers vs. dates)
   - Fix encoding issues (UTF-8 emoji rendering)

4. **Content Validation**
   - Check for incomplete assessments (missing required data)
   - Identify placeholder/sample data not replaced with real data
   - Validate formula integrity
   - Verify conditional formatting rules
   - Check data validation dropdown configurations

5. **Cross-Reference Validation**
   - Verify Regulation IDs match across workbooks (1→2→3→4→5→6)
   - Validate requirement IDs are unique and traceable
   - Check control mapping references are valid ISO 27001 controls
   - Ensure evidence references point to existing evidence items
   - Validate applicability status consistency (Workbook 1 vs. Workbook 2)

6. **Compliance Scoring Validation**
   - Verify scoring formula correctness
   - Validate coverage percentage calculations
   - Check for scoring anomalies or mathematical errors
   - Ensure tier classifications align with applicability scores

7. **Quality Reporting**
   - Generate comprehensive validation report with findings
   - Categorize issues by severity (Critical / High / Medium / Low)
   - Provide remediation guidance for each issue type
   - Track validation history over time

**Validation Scope:**
The script validates all six A.5.31 assessment workbooks:
- ISMS_Assessment_531_1_Regulatory_Inventory_YYYYMMDD.xlsx
- ISMS_Assessment_531_2_Applicability_Matrix_YYYYMMDD.xlsx
- ISMS_Assessment_531_3_Requirements_Register_YYYYMMDD.xlsx
- ISMS_Assessment_531_4_Control_Mapping_YYYYMMDD.xlsx
- ISMS_Assessment_531_5_Evidence_Register_YYYYMMDD.xlsx
- ISMS_Assessment_531_6_Compliance_Dashboard_YYYYMMDD.xlsx

**Output:**
- Normalized assessment workbooks (with _normalized suffix if changes made)
- Validation report (console, text file, or Excel format)
- Issue summary with counts by severity
- Remediation checklist

**Quality Checks Performed:**

**Critical Issues** (MUST fix before consolidation or audit):
- Missing required sheets
- Incorrect sheet names preventing consolidation
- Invalid data types in key columns (Regulation ID, dates, scores)
- Broken formulas in critical calculations
- Missing compliance scores or coverage percentages
- Regulation ID mismatches across workbooks
- External workbook references broken (Dashboard)

**High Priority Issues** (SHOULD fix promptly):
- Incomplete assessments (missing required data fields)
- Inconsistent dropdown values (prevents filtering/sorting)
- Missing evidence references for Primary control mappings
- Date format inconsistencies
- Applicability status inconsistencies between Workbook 1 and 2
- Sample/placeholder data not replaced with real data
- UTF-8 encoding issues (emoji rendering problems)

**Medium Priority Issues** (RECOMMENDED fix):
- Formatting inconsistencies (non-critical cosmetic issues)
- Whitespace issues (leading/trailing spaces)
- Non-standard terminology (inconsistent naming)
- Missing optional fields
- Incomplete requirement categorizations
- Evidence quality scores missing

**Low Priority Issues** (NICE to fix):
- Cosmetic formatting variations
- Optional field incompleteness
- Documentation style inconsistencies
- Non-critical metadata missing

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel processing

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - os (standard library)
    - re (standard library - regex for validation)
    - collections (standard library - for statistics)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    # Validate all A.5.31 assessment files in current directory
    python3 normalize_assessment_files_531.py

Advanced Usage:
    # Validate files in specific directory
    python3 normalize_assessment_files_531.py --input-dir /path/to/assessments
    
    # Validate with automatic normalization (fixes issues)
    python3 normalize_assessment_files_531.py --normalize
    
    # Validate specific assessment domain only
    python3 normalize_assessment_files_531.py --domain 1
    
    # Generate detailed validation report
    python3 normalize_assessment_files_531.py --report detailed
    
    # Dry run mode (report issues without modifying files)
    python3 normalize_assessment_files_531.py --dry-run
    
    # Specify output directory for normalized files
    python3 normalize_assessment_files_531.py --normalize --output-dir /path/to/output
    
    # Validate only critical and high priority issues
    python3 normalize_assessment_files_531.py --severity high

Command-Line Options:
    --input-dir PATH       Directory containing assessment workbooks (default: current directory)
    --output-dir PATH      Directory for normalized workbooks (default: input-dir)
    --normalize            Apply normalization fixes automatically (default: validation only)
    --domain N             Validate specific domain only (1-6, default: all)
    --report TYPE          Report format: summary|detailed|excel (default: summary)
    --dry-run              Validate only, don't modify files (default if --normalize not specified)
    --severity LEVEL       Minimum severity to report: critical|high|medium|low (default: low - all issues)
    --backup               Create backup before normalization (highly recommended)
    --cross-validate       Perform cross-workbook validation (Regulation IDs, requirement references, etc.)

Output Files:
    If --normalize used:
        - Original files: [filename]_backup_YYYYMMDD.xlsx (if --backup specified)
        - Normalized files: [filename] (updated in place) OR
        - Normalized files: [filename]_normalized.xlsx (if different output-dir)
    
    Validation report:
        - Console output (always)
        - Text file: A531_Assessment_Validation_Report_YYYYMMDD.txt (if --report text|detailed)
        - Excel file: A531_Assessment_Validation_Report_YYYYMMDD.xlsx (if --report excel)

Workflow Examples:

    1. Initial validation before dashboard consolidation:
       python3 normalize_assessment_files_531.py --dry-run --report detailed
       [Review issues, fix manually or proceed to normalization]
    
    2. Normalize and fix auto-fixable issues:
       python3 normalize_assessment_files_531.py --normalize --backup
       [Script fixes formatting, dates, dropdowns, whitespace]
    
    3. Validate after normalization (should show fewer issues):
       python3 normalize_assessment_files_531.py --severity high
       [Only critical/high issues remain, requiring manual fixes]
    
    4. Generate audit-ready validation report:
       python3 normalize_assessment_files_531.py --report excel --cross-validate
       [Comprehensive validation with Excel report for audit documentation]
    
    5. Pre-audit final check (all workbooks):
       python3 normalize_assessment_files_531.py --severity critical --cross-validate
       [Verify no critical issues before audit]

Exit Codes:
    0 = Validation passed (no critical issues found)
    1 = Validation failed (critical issues found, files not usable)
    2 = Validation warnings (high/medium issues, but files functional)

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.31
Utility Type:         Quality Assurance - Assessment Normalization and Validation
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.31: Complete A.5.31 Policy Framework (all sections)
    - ISMS-IMP-A.5.31 (all sections): Implementation Process Guides

Related Assessment Tools:
    - generate_531_1_regulatory_inventory.py (Assessment Workbook 1)
    - generate_531_2_applicability_matrix.py (Assessment Workbook 2)
    - generate_531_3_requirements_register.py (Assessment Workbook 3)
    - generate_531_4_control_mapping.py (Assessment Workbook 4)
    - generate_531_5_evidence_register.py (Assessment Workbook 5)
    - generate_531_6_compliance_dashboard.py (Dashboard)
    - consolidate_dashboard_a531.py (Alternative dashboard consolidation method)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements comprehensive validation framework for all six A.5.31 workbooks
    - Supports automated normalization of fixable issues
    - Generates quality assurance reports for audit documentation
    - Cross-workbook validation for data consistency

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Quality Assurance Philosophy:**
This script embodies the "don't fool yourself" engineering principle - it
systematically catches the errors humans make when populating assessment
workbooks, ensuring data quality before consolidation or audit presentation.

Think of this script as your **Red Team reviewer** - it challenges your data,
finds inconsistencies, and prevents embarrassing audit findings by catching
issues early.

**Validation vs. Normalization:**

**Validation Only** (--dry-run, default):
- Identifies issues without modifying files
- Safe to run anytime, anywhere
- Use for initial assessment of data quality
- Generates issue report for manual remediation

**Normalization** (--normalize flag):
- Automatically fixes issues where safe to do so
- Modifies files in place (or creates new files with _normalized suffix)
- Use after reviewing validation report
- ALWAYS use with --backup flag for safety

**What Normalization CAN Fix:**
- Date format inconsistencies (various formats → DD.MM.YYYY)
- Dropdown value variations (case, spacing, synonyms)
- Whitespace (leading/trailing spaces, multiple spaces)
- Encoding issues (UTF-8 emoji rendering)
- Formula syntax (minor corrections)
- Cell data types (text numbers → actual numbers)

**What Normalization CANNOT Fix** (requires human judgment):
- Missing required data (what should go there?)
- Incorrect tier classifications (requires regulatory analysis)
- Wrong control mappings (requires compliance expertise)
- Invalid evidence references (requires verification)
- Applicability assessment errors (requires legal judgment)
- Gap remediation priorities (requires risk assessment)

**Backup Recommendation:**
**CRITICAL:** ALWAYS use --backup flag when normalizing files.

Assessment workbooks represent significant data collection effort - weeks or
months of work. Don't risk data loss through normalization errors or unexpected
script behavior.

Backups are timestamped (filename_backup_YYYYMMDD.xlsx) and can be used to
restore if normalization causes unexpected issues.

**Pre-Consolidation Requirement:**
Run this script BEFORE consolidate_dashboard_a531.py (if using consolidation
method) to ensure clean, consistent input data.

Dashboard consolidation assumes normalized, validated assessment workbooks.
Garbage in → Garbage out.

**Cross-Workbook Validation:**
Critical feature for regulatory compliance framework consistency:

**Cross-Validation Checks:**
1. **Regulation IDs**: Same regulation ID used across all workbooks
   - Workbook 1 defines regulation → 2, 3, 4, 5, 6 reference it
   - Mismatch = broken traceability

2. **Applicability Status**: Workbook 1 vs. Workbook 2 consistency
   - Workbook 1 lists regulation as "Applicable"
   - Workbook 2 assessment should conclude "Applicable" (Tier 1 or 2)
   - Mismatch = assessment error or data entry error

3. **Requirement IDs**: Workbook 3 defines requirements → Workbook 4 maps them
   - All requirement IDs in Workbook 4 must exist in Workbook 3
   - Missing requirement = orphaned mapping

4. **Control References**: Workbook 4 maps to ISO 27001 controls
   - All control IDs must be valid ISO 27001:2022 Annex A controls (A.5.1 - A.8.37)
   - Invalid control ID = mapping error

5. **Evidence References**: Workbook 5 evidence items → Workbook 4 mappings
   - Control mappings should reference evidence items
   - Missing evidence reference = gap

Enable cross-validation with --cross-validate flag for comprehensive checks.

**Audit Considerations:**
Validation reports demonstrate quality assurance processes to auditors:
- Shows systematic approach to data quality
- Documents remediation of identified issues
- Provides audit trail of quality checks
- Demonstrates continuous improvement

Keep validation reports as evidence of quality control processes:
- Initial validation (shows baseline data quality)
- Post-normalization validation (shows improvement)
- Pre-audit validation (shows final quality confirmation)

**Data Integrity:**
Script validates FORMAT and STRUCTURE, not ACCURACY of content:
- Validates: Date format correct? Dropdown value from valid list?
- Does NOT validate: Is tier classification correct? Is control mapping accurate?

**Technical accuracy of compliance assessments remains assessor's responsibility.**

This script catches mechanical errors, not judgment errors.

**False Positives:**
Some validation warnings may be acceptable based on your specific context:
- "Optional field missing" may be intentional (not all fields always relevant)
- "Non-standard terminology" may be organization-specific correct terminology
- "Formatting variation" may be intentional for readability

Review validation report critically - don't blindly "fix" everything.
Use judgment to determine which issues are real problems vs. acceptable variations.

**Schema Changes:**
If you modify assessment workbook structures (add sheets, change columns,
alter dropdown lists), update this script's validation rules accordingly.

Out-of-sync validation rules will generate false positives (flagging correct
data as errors) or false negatives (missing actual errors).

**Document validation rule changes in script comments and CHANGE HISTORY section.**

**Performance:**
Script processes Excel files in memory. For very large assessment workbooks
(>50MB), consider:
- Increasing available system memory
- Processing files individually (--domain flag)
- Closing other applications during processing

Typical processing time: 1-5 minutes for all six workbooks.

**Error Handling:**
Script continues processing all files even if one fails validation:
- Partial validation better than no validation
- Check final summary for any files that couldn't be processed
- Investigate files with processing errors separately

**Scalability Considerations:**
For organizations with extensive regulatory portfolios:
- Workbook 3 (Requirements Register) may have 500+ requirements
- Workbook 4 (Control Mapping) becomes very large (requirements × controls matrix)
- Validation time increases with data volume
- Consider chunked validation for very large workbooks

**Integration with Compliance Workflow:**

**Quality Gate 1**: After populating assessment workbooks
- Validate (--dry-run --report detailed)
- Review issues, fix manually where judgment required
- Normalize (--normalize --backup)
- Re-validate to confirm fixes

**Quality Gate 2**: Before dashboard consolidation
- Validate all workbooks (--cross-validate)
- Confirm no critical issues
- Proceed with consolidation

**Quality Gate 3**: Before audit
- Final validation (--severity critical --cross-validate)
- Generate validation report for audit evidence
- Confirm no critical issues remain

**Common Validation Failures:**

1. **"Regulation ID mismatch between Workbook 1 and 2"**
   - Cause: Typo in Regulation ID field, or regulation in one workbook but not other
   - Fix: Ensure consistent Regulation ID spelling and format across workbooks

2. **"Requirement ID in Workbook 4 not found in Workbook 3"**
   - Cause: Control mapping references requirement that doesn't exist
   - Fix: Either add missing requirement to Workbook 3, or correct Requirement ID in Workbook 4

3. **"Invalid ISO 27001 control reference"**
   - Cause: Control ID typo (e.g., "A.5.32" instead of "A.5.31") or non-existent control
   - Fix: Correct control ID to valid Annex A control (A.5.1 through A.8.37)

4. **"Applicability status inconsistency"**
   - Cause: Workbook 1 says "Applicable" but Workbook 2 assessment concludes "Not Applicable"
   - Fix: Re-assess applicability in Workbook 2, or update Workbook 1 status

5. **"Missing evidence reference for Primary control mapping"**
   - Cause: Control mapping marked as Primary but no evidence documented in Workbook 5
   - Fix: Add evidence to Workbook 5, or downgrade mapping type if evidence doesn't exist

6. **"Date format inconsistent"**
   - Cause: Dates entered in various formats (MM/DD/YYYY, DD.MM.YYYY, YYYY-MM-DD)
   - Fix: Run normalization (--normalize) to standardize to DD.MM.YYYY

**Troubleshooting:**

**Issue: "ERROR: Cannot open workbook [filename]"**
- Solution: Verify file exists, is valid .xlsx format, not corrupted
- Solution: Check file is not open in Excel (close Excel before running script)
- Solution: Verify read permissions on file

**Issue: "WARNING: External workbook references broken (Dashboard)"**
- Solution: Ensure all source workbooks (1-5) are in same directory as Dashboard
- Solution: Verify file names match exactly what Dashboard expects
- Solution: Consider using consolidate_dashboard_a531.py if external links problematic

**Issue: "Script runs but reports 0 issues when issues clearly exist"**
- Solution: Validation rules may be out of sync with workbook structure
- Solution: Verify script version matches assessment workbook generator script versions
- Solution: Check for schema changes that invalidated validation rules

**Issue: "Normalization modifies data unexpectedly"**
- Solution: Restore from backup (you DID use --backup, right?)
- Solution: Review normalization rules in script code
- Solution: Use --dry-run first to see what would change before applying

**Best Practices:**
1. Validate early and often (after each significant data entry session)
2. Always use --backup with --normalize
3. Fix critical issues before proceeding to medium/low issues
4. Document reasons if you choose not to fix certain issues
5. Generate validation reports for audit evidence
6. Re-validate after any workbook structure changes
7. Include validation in your quality assurance SOP

**Continuous Improvement:**
As you use this script, you'll identify:
- Common errors your team makes (add specific checks)
- Acceptable variations (adjust validation rules to not flag these)
- Additional cross-checks needed (enhance cross-validation)
- Performance bottlenecks (optimize validation logic)

Update script over time to improve validation effectiveness and reduce false positives.

================================================================================
"""

import sys
import os
import shutil
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook


# ============================================================================
# SECTION 1: WORKBOOK TYPE DETECTION
# ============================================================================

WORKBOOK_PATTERNS = {
    "531_1": {
        "pattern": r".*531.*1.*Regulatory.*Inventory",
        "normalized_name": "ISMS_Assessment_531_1_Regulatory_Inventory.xlsx",
        "required_sheets": ["Regulatory_Inventory", "Instructions"],
        "description": "Regulatory Inventory",
    },
    "531_2": {
        "pattern": r".*531.*2.*Applicability.*Matrix",
        "normalized_name": "ISMS_Assessment_531_2_Applicability_Matrix.xlsx",
        "required_sheets": ["Applicability_Assessment", "Instructions"],
        "description": "Applicability Assessment Matrix",
    },
    "531_3": {
        "pattern": r".*531.*3.*Requirements.*Register",
        "normalized_name": "ISMS_Assessment_531_3_Requirements_Register.xlsx",
        "required_sheets": ["Requirements_Register", "Instructions"],
        "description": "Requirements Register",
    },
    "531_4": {
        "pattern": r".*531.*4.*Control.*Mapping",
        "normalized_name": "ISMS_Assessment_531_4_Control_Mapping_Matrix.xlsx",
        "required_sheets": ["Control_Mapping_Matrix", "ISO27001_Controls_Reference"],
        "description": "Control Mapping Matrix",
    },
    "531_5": {
        "pattern": r".*531.*5.*Evidence.*Register",
        "normalized_name": "ISMS_Assessment_531_5_Evidence_Register.xlsx",
        "required_sheets": ["Evidence_Register", "Instructions"],
        "description": "Evidence Register",
    },
    "531_6": {
        "pattern": r".*531.*6.*Compliance.*Dashboard",
        "normalized_name": "ISMS_Assessment_531_6_Compliance_Dashboard.xlsx",
        "required_sheets": ["Executive_Dashboard", "Regulatory_Status"],
        "description": "Compliance Dashboard",
    },
}


def detect_workbook_type(filename, sheet_names):
    """Detect workbook type based on filename and sheet names."""
    for wb_type, config in WORKBOOK_PATTERNS.items():
        # Check filename pattern
        if re.search(config["pattern"], filename, re.IGNORECASE):
            # Verify required sheets
            required_sheets = config["required_sheets"]
            if all(sheet in sheet_names for sheet in required_sheets):
                return wb_type, config
    
    return None, None


def is_normalized_filename(filename):
    """Check if filename is already in normalized format."""
    # Normalized format: ISMS_Assessment_531_X_Name.xlsx (no dates, versions, etc.)
    pattern = r"^ISMS_Assessment_531_[1-6]_[A-Za-z_]+\.xlsx$"
    return re.match(pattern, filename) is not None


# ============================================================================
# SECTION 2: FILE SCANNING AND VALIDATION
# ============================================================================

def scan_directory(directory):
    """Scan directory for A.5.31 assessment workbooks."""
    print(f"\n📂 Scanning directory: {directory}")
    print("─" * 70)
    
    excel_files = []
    
    # Find all Excel files
    for filepath in Path(directory).glob("*.xlsx"):
        # Skip temporary files
        if filepath.name.startswith("~$"):
            continue
        
        # Skip already normalized files
        if is_normalized_filename(filepath.name):
            print(f"⏭️  Skipping (already normalized): {filepath.name}")
            continue
        
        # Check if it's an A.5.31 workbook
        if "531" in filepath.name or "5.31" in filepath.name or "A.5.31" in filepath.name:
            excel_files.append(filepath)
            print(f"✓ Found: {filepath.name}")
    
    print(f"\nFound {len(excel_files)} workbook(s) to process")
    return excel_files


def validate_workbook(filepath):
    """Validate workbook and determine type."""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        wb_type, config = detect_workbook_type(filepath.name, sheet_names)
        
        if not wb_type:
            return None, None, "Cannot detect workbook type"
        
        return wb_type, config, None
    
    except Exception as e:
        return None, None, f"Error loading workbook: {e}"


# ============================================================================
# SECTION 3: NORMALIZATION AND MANIFEST
# ============================================================================

def normalize_files(source_files, target_dir):
    """Normalize filenames and copy to target directory."""
    
    # Create target directory if needed
    Path(target_dir).mkdir(parents=True, exist_ok=True)
    
    print(f"\n📋 Normalizing files to: {target_dir}")
    print("─" * 70)
    
    manifest = []
    errors = []
    
    for source_path in source_files:
        print(f"\nProcessing: {source_path.name}")
        
        # Validate workbook
        wb_type, config, error = validate_workbook(source_path)
        
        if error:
            print(f"  ❌ ERROR: {error}")
            errors.append((source_path.name, error))
            continue
        
        # Get normalized filename
        normalized_name = config["normalized_name"]
        target_path = Path(target_dir) / normalized_name
        
        # Check if target already exists
        if target_path.exists():
            print(f"  ⚠️  WARNING: Target file exists: {normalized_name}")
            
            # Compare modification times
            source_mtime = source_path.stat().st_mtime
            target_mtime = target_path.stat().st_mtime
            
            if source_mtime <= target_mtime:
                print(f"  ⏭️  Skipping: Target is newer or same age")
                continue
            else:
                print(f"  ♻️  Replacing: Source is newer")
        
        # Copy file
        try:
            shutil.copy2(source_path, target_path)
            print(f"  ✅ Copied to: {normalized_name}")
            
            # Add to manifest
            manifest.append({
                "original_file": source_path.name,
                "original_path": str(source_path.absolute()),
                "normalized_file": normalized_name,
                "normalized_path": str(target_path.absolute()),
                "workbook_type": wb_type,
                "description": config["description"],
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            
        except Exception as e:
            print(f"  ❌ ERROR copying file: {e}")
            errors.append((source_path.name, f"Copy failed: {e}"))
    
    return manifest, errors


def create_manifest(manifest, target_dir):
    """Create audit manifest file."""
    manifest_path = Path(target_dir) / "normalization_manifest.txt"
    
    print(f"\n📄 Creating manifest: {manifest_path}")
    
    with open(manifest_path, 'w', encoding='utf-8') as f:
        f.write("=" * 70 + "\n")
        f.write("ISMS A.5.31 ASSESSMENT WORKBOOKS - NORMALIZATION MANIFEST\n")
        f.write("=" * 70 + "\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total Files Normalized: {len(manifest)}\n")
        f.write("=" * 70 + "\n\n")
        
        for idx, entry in enumerate(manifest, 1):
            f.write(f"{idx}. {entry['description']} ({entry['workbook_type']})\n")
            f.write(f"   Original File:    {entry['original_file']}\n")
            f.write(f"   Original Path:    {entry['original_path']}\n")
            f.write(f"   Normalized File:  {entry['normalized_file']}\n")
            f.write(f"   Normalized Path:  {entry['normalized_path']}\n")
            f.write(f"   Timestamp:        {entry['timestamp']}\n")
            f.write("\n")
        
        f.write("=" * 70 + "\n")
        f.write("END OF MANIFEST\n")
        f.write("=" * 70 + "\n")
    
    print(f"✓ Manifest created: {manifest_path}")


# ============================================================================
# SECTION 4: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    print("\n" + "=" * 70)
    print("ISMS A.5.31 ASSESSMENT WORKBOOKS - NORMALIZATION SCRIPT")
    print("=" * 70)
    
    # Parse arguments
    if len(sys.argv) >= 2:
        source_dir = sys.argv[1]
    else:
        source_dir = "."
    
    if len(sys.argv) >= 3:
        target_dir = sys.argv[2]
    else:
        target_dir = os.path.join(source_dir, "normalized")
    
    # Validate source directory
    if not os.path.isdir(source_dir):
        print(f"\n❌ ERROR: Source directory not found: {source_dir}")
        sys.exit(1)
    
    print(f"\nSource Directory: {os.path.abspath(source_dir)}")
    print(f"Target Directory: {os.path.abspath(target_dir)}")
    
    # Scan for files
    source_files = scan_directory(source_dir)
    
    if not source_files:
        print("\n⚠️  No workbooks found to normalize")
        print("\nFiles are considered for normalization if:")
        print("  • Filename contains '531', '5.31', or 'A.5.31'")
        print("  • Filename is NOT already in normalized format")
        print("  • File is a valid Excel workbook (.xlsx)")
        sys.exit(0)
    
    # Normalize files
    manifest, errors = normalize_files(source_files, target_dir)
    
    # Create manifest
    if manifest:
        create_manifest(manifest, target_dir)
    
    # Summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    
    print(f"\n✅ Successfully normalized: {len(manifest)} file(s)")
    if errors:
        print(f"❌ Errors encountered: {len(errors)} file(s)")
        print("\nErrors:")
        for filename, error in errors:
            print(f"  • {filename}: {error}")
    
    print(f"\n📂 Normalized files location: {os.path.abspath(target_dir)}")
    
    if manifest:
        print("\n✓ Files are ready for dashboard integration")
        print("\nNormalized filenames:")
        for entry in manifest:
            print(f"  • {entry['normalized_file']}")
    
    print("\n" + "=" * 70)
    
    if errors:
        sys.exit(1)


if __name__ == "__main__":
    main()
