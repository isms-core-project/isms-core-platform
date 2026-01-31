#!/usr/bin/env python3
"""
Excel Sanity Check for ISMS A.5.31 Assessment Workbooks
Validates structure, data integrity, and quality of assessment workbooks.

Usage:
    python3 excel_sanity_check_531.py <workbook_file>
    python3 excel_sanity_check_531.py --all   # Check all workbooks in current directory
"""

import sys
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


# ============================================================================
# SECTION 1: WORKBOOK TYPE DETECTION
# ============================================================================

WORKBOOK_TYPES = {
    "531_1": {
        "name": "Regulatory Inventory",
        "required_sheets": ["Regulatory_Inventory", "Instructions", "Summary_Metrics"],
        "key_columns": ["Regulation ID", "Regulation Name", "Tier", "Applicability Status"],
        "validations": ["Tier", "Applicability Status"],
    },
    "531_2": {
        "name": "Applicability Assessment Matrix",
        "required_sheets": ["Applicability_Assessment", "Instructions"],
        "key_sections": ["SECTION A:", "SECTION B:", "SECTION E:"],
        "validations": ["Y/N"],
    },
    "531_3": {
        "name": "Requirements Register",
        "required_sheets": ["Requirements_Register", "Instructions", "Summary_Metrics"],
        "key_columns": ["Requirement ID", "Regulation ID", "Category", "Priority", "Implementation Status", "Gap Status"],
        "validations": ["Category", "Priority", "Implementation Status", "Gap Status"],
    },
    "531_4": {
        "name": "Control Mapping Matrix",
        "required_sheets": ["Control_Mapping_Matrix", "ISO27001_Controls_Reference"],
        "key_columns": ["Requirement ID", "Interpreted Requirement"],
        "validations": ["P/S/Su"],
    },
    "531_5": {
        "name": "Evidence Register",
        "required_sheets": ["Evidence_Register", "Instructions", "Evidence_by_Regulation", "Evidence_by_Control"],
        "key_columns": ["Evidence ID", "Requirement ID", "Control ID", "Evidence Type", "Verification Status"],
        "validations": ["Evidence Type", "Verification Status", "Refresh Frequency", "Audit Ready"],
    },
    "531_6": {
        "name": "Compliance Dashboard",
        "required_sheets": ["Executive_Dashboard", "Regulatory_Status", "Requirements_Progress", "Control_Coverage"],
        "key_sections": ["KEY PERFORMANCE INDICATORS", "CRITICAL ALERTS"],
    },
}


def detect_workbook_type(wb, filename):
    """Detect which type of A.5.31 workbook this is."""
    sheet_names = wb.sheetnames
    
    # Try to match based on filename first
    for wb_type, config in WORKBOOK_TYPES.items():
        if wb_type.replace("_", "-") in filename or wb_type in filename:
            return wb_type, config
    
    # Match based on sheet names
    for wb_type, config in WORKBOOK_TYPES.items():
        required_sheets = config.get("required_sheets", [])
        if all(sheet in sheet_names for sheet in required_sheets[:2]):  # Check first 2 required sheets
            return wb_type, config
    
    return None, None


# ============================================================================
# SECTION 2: VALIDATION FUNCTIONS
# ============================================================================

def check_required_sheets(wb, config):
    """Check if all required sheets exist."""
    issues = []
    warnings = []
    
    required_sheets = config.get("required_sheets", [])
    sheet_names = wb.sheetnames
    
    for sheet in required_sheets:
        if sheet not in sheet_names:
            issues.append(f"❌ MISSING REQUIRED SHEET: '{sheet}'")
        else:
            warnings.append(f"✓ Sheet '{sheet}' exists")
    
    return issues, warnings


def check_sheet_structure(wb, config):
    """Check sheet structure (columns, headers)."""
    issues = []
    warnings = []
    
    key_columns = config.get("key_columns", [])
    if not key_columns:
        return issues, warnings
    
    # Check first data sheet
    required_sheets = config.get("required_sheets", [])
    if not required_sheets:
        return issues, warnings
    
    first_sheet_name = required_sheets[0]
    if first_sheet_name not in wb.sheetnames:
        return issues, warnings
    
    ws = wb[first_sheet_name]
    
    # Get headers from row 1
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        if cell_value:
            headers.append(str(cell_value).strip())
    
    # Check for key columns
    missing_columns = []
    for key_col in key_columns:
        if key_col not in headers:
            missing_columns.append(key_col)
    
    if missing_columns:
        issues.append(f"❌ MISSING KEY COLUMNS in '{first_sheet_name}': {', '.join(missing_columns)}")
    else:
        warnings.append(f"✓ All key columns present in '{first_sheet_name}'")
    
    # Check if sheet has data (at least 2 rows: header + 1 data row)
    if ws.max_row < 2:
        warnings.append(f"⚠️  Sheet '{first_sheet_name}' has no data rows (only header)")
    else:
        warnings.append(f"✓ Sheet '{first_sheet_name}' has {ws.max_row - 1} data rows")
    
    return issues, warnings


def check_data_validations(wb, config):
    """Check if data validation rules exist."""
    issues = []
    warnings = []
    
    validations = config.get("validations", [])
    if not validations:
        return issues, warnings
    
    required_sheets = config.get("required_sheets", [])
    if not required_sheets:
        return issues, warnings
    
    first_sheet_name = required_sheets[0]
    if first_sheet_name not in wb.sheetnames:
        return issues, warnings
    
    ws = wb[first_sheet_name]
    
    # Check if sheet has any data validations
    if hasattr(ws, 'data_validations') and ws.data_validations.dataValidation:
        val_count = len(ws.data_validations.dataValidation)
        warnings.append(f"✓ Sheet '{first_sheet_name}' has {val_count} data validation rule(s)")
    else:
        warnings.append(f"⚠️  Sheet '{first_sheet_name}' has no data validation rules")
    
    return issues, warnings


def check_conditional_formatting(wb, config):
    """Check if conditional formatting exists."""
    issues = []
    warnings = []
    
    required_sheets = config.get("required_sheets", [])
    if not required_sheets:
        return issues, warnings
    
    first_sheet_name = required_sheets[0]
    if first_sheet_name not in wb.sheetnames:
        return issues, warnings
    
    ws = wb[first_sheet_name]
    
    # Check if sheet has conditional formatting
    if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting._cf_rules:
        cf_count = len(ws.conditional_formatting._cf_rules)
        warnings.append(f"✓ Sheet '{first_sheet_name}' has {cf_count} conditional formatting rule(s)")
    else:
        warnings.append(f"⚠️  Sheet '{first_sheet_name}' has no conditional formatting")
    
    return issues, warnings


def check_formulas(wb, config):
    """Check for formulas in summary/metrics sheets."""
    issues = []
    warnings = []
    
    # Look for Summary or Metrics sheet
    summary_sheets = ["Summary_Metrics", "Summary", "Metrics", "Gap_Summary"]
    
    found_formulas = False
    for sheet_name in summary_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            formula_count = 0
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
            
            if formula_count > 0:
                warnings.append(f"✓ Sheet '{sheet_name}' has {formula_count} formula(s)")
                found_formulas = True
            else:
                warnings.append(f"⚠️  Sheet '{sheet_name}' has no formulas")
    
    if not found_formulas:
        warnings.append(f"ℹ️  No summary sheets with formulas found")
    
    return issues, warnings


def check_utf8_encoding(wb, config):
    """Check for proper UTF-8 encoding (emoji support)."""
    issues = []
    warnings = []
    
    required_sheets = config.get("required_sheets", [])
    if not required_sheets:
        return issues, warnings
    
    first_sheet_name = required_sheets[0]
    if first_sheet_name not in wb.sheetnames:
        return issues, warnings
    
    ws = wb[first_sheet_name]
    
    # Check for emoji characters in data
    emoji_found = False
    emoji_chars = ['✅', '❌', '⚠️', '⏳', '🔴', '🟡', '🟢', '📄', '📋', '⚙️', '📊', '📈', '🎓', '🧪', '📸']
    
    for row in ws.iter_rows(max_row=100):  # Check first 100 rows
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if any(emoji in cell.value for emoji in emoji_chars):
                    emoji_found = True
                    break
        if emoji_found:
            break
    
    if emoji_found:
        warnings.append(f"✓ UTF-8 encoding detected (emoji characters found)")
    else:
        warnings.append(f"ℹ️  No emoji characters found (may be blank template)")
    
    return issues, warnings


def check_freeze_panes(wb, config):
    """Check if freeze panes are set."""
    issues = []
    warnings = []
    
    required_sheets = config.get("required_sheets", [])
    if not required_sheets:
        return issues, warnings
    
    first_sheet_name = required_sheets[0]
    if first_sheet_name not in wb.sheetnames:
        return issues, warnings
    
    ws = wb[first_sheet_name]
    
    if ws.freeze_panes:
        warnings.append(f"✓ Freeze panes set at: {ws.freeze_panes}")
    else:
        warnings.append(f"⚠️  No freeze panes set")
    
    return issues, warnings


def check_auto_filter(wb, config):
    """Check if auto-filter is enabled."""
    issues = []
    warnings = []
    
    required_sheets = config.get("required_sheets", [])
    if not required_sheets:
        return issues, warnings
    
    first_sheet_name = required_sheets[0]
    if first_sheet_name not in wb.sheetnames:
        return issues, warnings
    
    ws = wb[first_sheet_name]
    
    if ws.auto_filter and ws.auto_filter.ref:
        warnings.append(f"✓ Auto-filter enabled on: {ws.auto_filter.ref}")
    else:
        warnings.append(f"⚠️  No auto-filter enabled")
    
    return issues, warnings


# ============================================================================
# SECTION 3: MAIN SANITY CHECK FUNCTION
# ============================================================================

def sanity_check_workbook(filepath):
    """Perform complete sanity check on a workbook."""
    print("\n" + "=" * 70)
    print(f"SANITY CHECK: {os.path.basename(filepath)}")
    print("=" * 70)
    
    # Try to load workbook
    try:
        wb = load_workbook(filepath, data_only=False)
    except InvalidFileException as e:
        print(f"\n❌ CRITICAL ERROR: Cannot open file - {e}")
        return False
    except Exception as e:
        print(f"\n❌ CRITICAL ERROR: {e}")
        return False
    
    # Detect workbook type
    wb_type, config = detect_workbook_type(wb, os.path.basename(filepath))
    
    if not wb_type:
        print("\n⚠️  WARNING: Cannot detect workbook type")
        print("   File may not be an A.5.31 assessment workbook")
        return False
    
    print(f"\n📋 Workbook Type: {config['name']} ({wb_type})")
    
    # Run all checks
    all_issues = []
    all_warnings = []
    
    checks = [
        ("Required Sheets", check_required_sheets),
        ("Sheet Structure", check_sheet_structure),
        ("Data Validations", check_data_validations),
        ("Conditional Formatting", check_conditional_formatting),
        ("Formulas", check_formulas),
        ("UTF-8 Encoding", check_utf8_encoding),
        ("Freeze Panes", check_freeze_panes),
        ("Auto Filter", check_auto_filter),
    ]
    
    print("\n" + "─" * 70)
    print("RUNNING CHECKS...")
    print("─" * 70)
    
    for check_name, check_func in checks:
        print(f"\n{check_name}:")
        issues, warnings = check_func(wb, config)
        
        for issue in issues:
            print(f"  {issue}")
            all_issues.append(issue)
        
        for warning in warnings:
            print(f"  {warning}")
            all_warnings.append(warning)
    
    # Summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    
    if all_issues:
        print(f"\n❌ ISSUES FOUND: {len(all_issues)}")
        for issue in all_issues:
            print(f"   {issue}")
        print("\n⚠️  WORKBOOK HAS ISSUES - NEEDS ATTENTION")
        return False
    else:
        print(f"\n✅ NO CRITICAL ISSUES FOUND")
        print(f"ℹ️  Informational items: {len(all_warnings)}")
        print("\n✓ WORKBOOK PASSED SANITY CHECK")
        return True


# ============================================================================
# SECTION 4: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    if len(sys.argv) < 2:
        print("Usage: python3 excel_sanity_check_531.py <workbook_file>")
        print("       python3 excel_sanity_check_531.py --all")
        sys.exit(1)
    
    if sys.argv[1] == "--all":
        # Check all Excel files in current directory
        excel_files = list(Path(".").glob("ISMS_Assessment_531_*.xlsx"))
        
        if not excel_files:
            print("No A.5.31 assessment workbooks found in current directory")
            sys.exit(1)
        
        print(f"\nFound {len(excel_files)} workbook(s) to check\n")
        
        results = {}
        for filepath in sorted(excel_files):
            passed = sanity_check_workbook(str(filepath))
            results[filepath.name] = passed
        
        # Overall summary
        print("\n" + "=" * 70)
        print("OVERALL SUMMARY")
        print("=" * 70)
        
        passed_count = sum(1 for p in results.values() if p)
        failed_count = len(results) - passed_count
        
        for filename, passed in results.items():
            status = "✅ PASS" if passed else "❌ FAIL"
            print(f"{status}: {filename}")
        
        print(f"\nTotal: {len(results)} workbooks")
        print(f"Passed: {passed_count}")
        print(f"Failed: {failed_count}")
        
        if failed_count > 0:
            sys.exit(1)
    else:
        # Check single file
        filepath = sys.argv[1]
        
        if not os.path.exists(filepath):
            print(f"Error: File not found: {filepath}")
            sys.exit(1)
        
        passed = sanity_check_workbook(filepath)
        
        if not passed:
            sys.exit(1)


if __name__ == "__main__":
    main()
