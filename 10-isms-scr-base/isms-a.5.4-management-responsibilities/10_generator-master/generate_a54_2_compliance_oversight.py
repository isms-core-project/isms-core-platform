#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.4.2 - Compliance Oversight Tracker
================================================================================

ISO/IEC 27001:2022 Control A.5.4: Management Responsibilities
Assessment Domain 2 of 3: Compliance Oversight

This script generates an Excel workbook for tracking how managers ensure
their teams comply with information security requirements.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.4.2"
WORKBOOK_NAME = "Compliance Oversight Tracker"
CONTROL_ID = "A.5.4"
CONTROL_NAME = "Management Responsibilities"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions sheet."""
    ws.title = "Instructions"

    instructions = [
        ["ISMS-IMP-A.5.4.2 - Compliance Oversight Tracker"],
        [""],
        ["PURPOSE"],
        ["This workbook tracks management oversight of team security compliance"],
        ["per ISMS-POL-A.5.4 (Management Responsibilities)."],
        [""],
        ["SHEETS"],
        ["1. Instructions - This guidance sheet"],
        ["2. Training Oversight - Manager tracking of team training completion"],
        ["3. Policy Violations - Manager handling of policy violations"],
        ["4. Access Reviews - Manager participation in access reviews"],
        ["5. Quarterly Summary - Aggregated oversight metrics"],
        [""],
        ["WORKFLOW"],
        ["1. Track training completion oversight per manager"],
        ["2. Log policy violations and manager response"],
        ["3. Record access review participation"],
        ["4. Generate quarterly summary for management review"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SHEETS", "WORKFLOW"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 70


def create_training_oversight_sheet(ws):
    """Create the Training Oversight sheet."""
    ws.title = "Training Oversight"

    headers = [
        "Manager_ID", "Manager_Name", "Department", "Team_Size",
        "Training_Required", "Training_Completed", "Completion_Rate",
        "Overdue_Count", "Follow_Up_Actions", "Last_Review_Date"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Template row with formulas
    row = 2
    ws.cell(row=row, column=7, value=f"=IF(E{row}>0,F{row}/E{row}*100,0)")

    # Format input rows
    for row in range(2, 51):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            if col != 7:  # Not formula column
                cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 25, 20, 10, 15, 18, 15, 12, 35, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_policy_violations_sheet(ws):
    """Create the Policy Violations sheet."""
    ws.title = "Policy Violations"

    headers = [
        "Violation_ID", "Date_Reported", "Manager_ID", "Manager_Name",
        "Employee_ID", "Violation_Type", "Severity", "Manager_Response",
        "Response_Date", "Resolution", "Lessons_Learned"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(severity_dv)
    severity_dv.add('G2:G100')

    response_dv = DataValidation(type="list", formula1='"Coaching,Written Warning,Performance Plan,Escalated,No Action"')
    ws.add_data_validation(response_dv)
    response_dv.add('H2:H100')

    resolution_dv = DataValidation(type="list", formula1='"Resolved,In Progress,Escalated,Pending"')
    ws.add_data_validation(resolution_dv)
    resolution_dv.add('J2:J100')

    # Format input rows
    for row in range(2, 101):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 15, 12, 25, 12, 25, 10, 18, 15, 12, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_access_reviews_sheet(ws):
    """Create the Access Reviews sheet."""
    ws.title = "Access Reviews"

    headers = [
        "Review_ID", "Review_Period", "Manager_ID", "Manager_Name",
        "System_Scope", "Accounts_Reviewed", "Changes_Requested",
        "Review_Completed", "Completion_Date", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    completed_dv = DataValidation(type="list", formula1='"Yes,No,Partial"')
    ws.add_data_validation(completed_dv)
    completed_dv.add('H2:H100')

    # Format input rows
    for row in range(2, 101):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 15, 12, 25, 25, 18, 18, 15, 15, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_quarterly_summary_sheet(ws):
    """Create the Quarterly Summary sheet."""
    ws.title = "Quarterly Summary"

    headers = [
        "Manager_ID", "Manager_Name", "Quarter", "Training_Compliance_Rate",
        "Violations_Handled", "Avg_Response_Days", "Access_Reviews_Completed",
        "Overall_Rating", "Comments"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    quarter_dv = DataValidation(type="list", formula1='"Q1,Q2,Q3,Q4"')
    ws.add_data_validation(quarter_dv)
    quarter_dv.add('C2:C200')

    rating_dv = DataValidation(type="list", formula1='"Exceeds,Meets,Below,N/A"')
    ws.add_data_validation(rating_dv)
    rating_dv.add('H2:H200')

    # Format input rows
    for row in range(2, 201):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 25, 10, 22, 18, 18, 25, 15, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_workbook():
    """Generate the complete workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()
    default_sheet = wb.active

    create_instructions_sheet(wb.create_sheet())
    create_training_oversight_sheet(wb.create_sheet())
    create_policy_violations_sheet(wb.create_sheet())
    create_access_reviews_sheet(wb.create_sheet())
    create_quarterly_summary_sheet(wb.create_sheet())

    wb.remove(default_sheet)

    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - INITIAL CREATION
# QA_TOOL: Claude Code
# CHANGES: Initial generator creation for A.5.4 Management Responsibilities
# =============================================================================
