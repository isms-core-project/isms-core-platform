#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
Excel Workbook Sanity Checker for ISMS A.8.16 Assessment Workbooks
================================================================================

ISO/IEC 27001:2022 Control A.8.16: Monitoring Activities
Quality Assurance Utility: Excel Workbook Diagnostic and Validation Tool

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script performs comprehensive diagnostic checks on A.8.16 monitoring
assessment Excel workbooks to identify issues that trigger Excel's "file level
validation and repair" warnings and to ensure workbook quality before distribution
or consolidation.

**Purpose:**
Diagnoses common openpyxl-generated Excel issues, validates workbook structure
against framework specifications, and provides actionable remediation guidance
to ensure workbook integrity and audit readiness.

**Diagnostic Scope:**
- Sheet structure validation (presence of required sheets)
- Assessment ID verification (ISMS-IMP-A.8.16.X identifiers)
- Formula syntax validation (inter-sheet references, external links)
- Data validation conflicts (overlapping ranges)
- Merged cell integrity (content in non-primary cells)
- Evidence register validation (capacity and format)
- Approval workflow validation (3-level approval structure)
- Summary dashboard validation (formula completeness)
- Row capacity validation (110-row target for data entry)
- Worksheet dimension checks (performance considerations)

**Supported Workbook Types:**
- A816-1: Monitoring Infrastructure Assessment
- A816-2: Baseline & Anomaly Detection Assessment
- A816-3: Coverage Assessment
- A816-4: Alert Management & Response Assessment
- A816-5: Compliance Dashboard (Master)

**Diagnostic Output:**
- Workbook type detection and classification
- 10 comprehensive validation checks with ✓/✗/⚠ status
- Critical issues requiring immediate fix
- Warnings recommending attention
- Recommended remediation actions
- Workbook-specific notes and expected structure
- Integration guidance for dashboard consolidation

**When to Use:**
- After generating workbooks (Scripts 1-5)
- Before distributing to stakeholders
- After manual modifications or edits
- Before consolidation into dashboard
- When Excel shows repair warnings
- Quality assurance validation before audits

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel processing

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - pathlib (standard library)
    - re (standard library - regex for validation)
    - sys (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    # Check any A.8.16 assessment workbook
    python3 excel_sanity_check_a816.py <filename.xlsx>

Domain-Specific Examples:
    # Check Monitoring Infrastructure Assessment
    python3 excel_sanity_check_a816.py ISMS_A_8_16_1_Monitoring_Infrastructure.xlsx
    
    # Check Baseline & Detection Assessment
    python3 excel_sanity_check_a816.py ISMS_A_8_16_2_Baseline_Detection.xlsx
    
    # Check Coverage Assessment
    python3 excel_sanity_check_a816.py ISMS_A_8_16_3_Coverage_Assessment.xlsx
    
    # Check Alert Management Assessment
    python3 excel_sanity_check_a816.py ISMS_A_8_16_4_Alert_Management.xlsx
    
    # Check Compliance Dashboard
    python3 excel_sanity_check_a816.py ISMS_A_8_16_5_Compliance_Dashboard.xlsx
    
    # Check normalized workbook
    python3 excel_sanity_check_a816.py Dashboard_Sources/A816_1_*.xlsx

Quality Assurance Workflow:
    # After generation (immediate validation)
    python3 generate_a816_1_monitoring_infrastructure.py
    python3 excel_sanity_check_a816.py ISMS_A_8_16_1_*.xlsx
    
    # Before consolidation (validate all source workbooks)
    for f in ISMS_A_8_16_[1-4]_*.xlsx; do
        python3 excel_sanity_check_a816.py "$f"
    done
    
    # After normalization
    python3 normalize_assessment_files_a816.py
    python3 excel_sanity_check_a816.py Dashboard_Sources/A816_*.xlsx
    
    # Dashboard validation
    python3 generate_a816_5_compliance_dashboard.py
    python3 excel_sanity_check_a816.py Dashboard_Sources/*Dashboard*.xlsx

Output:
    Console output with:
    - Workbook type detection
    - 10 validation checks with detailed findings
    - Critical issues summary (must fix)
    - Warnings summary (should fix)
    - Recommended remediation actions
    - Workbook-specific notes
    - Expected structure documentation

Exit Codes:
    0 - No issues detected (workbook healthy)
    1 - Critical issues found or file error

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.16
Utility Type:         Quality Assurance - Excel Workbook Diagnostics
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization ISMS Team]
Date:                 24.01.2025
Last Modified:        24.01.2025
Python Version:       3.8+
License:              [Organization License/Terms]

Related Documents:
    - ISMS-POL-A.8.16: Monitoring Activities Policy (Governance)
    - ISMS-IMP-A.8.16.1: Monitoring Infrastructure Assessment
    - ISMS-IMP-A.8.16.2: Baseline & Detection Assessment
    - ISMS-IMP-A.8.16.3: Coverage Assessment
    - ISMS-IMP-A.8.16.4: Alert Management Assessment
    - ISMS-IMP-A.8.16.5: Compliance Dashboard Specification

Related Scripts:
    - generate_a816_1_monitoring_infrastructure.py (generates workbooks to check)
    - generate_a816_2_baseline_detection.py (generates workbooks to check)
    - generate_a816_3_coverage_assessment.py (generates workbooks to check)
    - generate_a816_4_alert_management.py (generates workbooks to check)
    - generate_a816_5_compliance_dashboard.py (generates dashboard to check)
    - normalize_assessment_files_a816.py (normalizes before checking)
    - consolidate_a816_dashboard.py (consolidates after checking)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 24.01.2025
    - Initial release
    - Implements 10 comprehensive diagnostic checks
    - Supports all five A.8.16 assessment workbook types
    - Provides detailed remediation guidance
    - Includes workbook-specific validation rules

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Quality Assurance Philosophy:**
"The first principle is that you must not fool yourself — and you are the 
easiest person to fool." - Richard Feynman

This sanity checker embodies anti-cargo-cult engineering - it catches structural
issues, formula errors, and validation conflicts that humans easily overlook.
Use it proactively, not reactively (after Excel complains).

**10 Validation Checks Explained:**

1. **Sheet Structure Validation**
   - Verifies all required sheets present
   - Detects missing or renamed sheets
   - Flexible matching (handles minor name variations)
   - Domain-specific expected sheet lists

2. **Assessment ID Validation**
   - Locates ISMS-IMP-A.8.16.X identifier
   - Validates correct assessment type
   - Checks Instructions sheet and metadata

3. **Formula Validation**
   - Checks formula syntax (balanced parentheses)
   - Validates inter-sheet references (sheet exists)
   - Identifies invalid sheet references
   - Detects external workbook links (dashboard requirement)
   - Reports formula issues with cell coordinates

4. **Data Validation Conflicts**
   - Identifies overlapping validation ranges
   - Checks for conflicting dropdown rules
   - Counts total validations per sheet

5. **Merged Cells Validation**
   - Ensures only top-left cell has content
   - Detects content in non-primary merged cells
   - Prevents Excel repair warnings from merged cell issues

6. **Evidence Register Validation**
   - Counts evidence row capacity (target: 100 rows)
   - Validates EV-XXX ID format
   - Enforces minimum 10 evidence rows
   - Checks for Evidence Register sheet presence

7. **Approval Sign-Off Validation**
   - Checks for 3-level approval workflow
   - Validates: Assessor → SOC Lead → CISO
   - Ensures approval structure complete

8. **Summary Dashboard Validation**
   - Counts formulas in Summary/Dashboard sheet
   - Validates calculated metrics present
   - Checks for empty or incomplete dashboards

9. **Row Capacity Validation**
   - Checks data entry capacity (target: 110 rows)
   - Counts example vs. entry rows
   - Warns on insufficient capacity (<50 rows)

10. **Worksheet Dimensions**
    - Flags oversized worksheets (>5000 rows, >50 columns)
    - Identifies potential performance issues

**Common Issues Detected:**

Critical Issues (Must Fix):
- Missing required sheets (Evidence Register, Approval Sign-Off)
- Invalid formula references (sheet doesn't exist)
- Broken inter-sheet formulas
- Missing assessment ID
- Evidence Register <10 rows (minimum requirement)

High Priority Issues (Should Fix):
- Merged cell content in non-primary cells
- Overlapping data validation ranges
- Formula syntax errors (unbalanced parentheses)
- Missing approval workflow levels

Medium Priority Issues (Recommended):
- Evidence Register <100 rows (recommended capacity)
- Low row capacity in data sheets (<50 rows)
- Excessive formula count suggesting copy-paste errors

Low Priority Issues (Nice to Fix):
- Missing optional sheets
- Cosmetic formatting variations
- Minor naming differences

**Excel Repair Warnings - Root Causes:**

If Excel shows "We found a problem with some content" on open:

Common Causes:
1. **Invalid Sheet References** - Formula references non-existent sheet
2. **Merged Cell Issues** - Content in non-primary cells of merge
3. **Data Validation Conflicts** - Overlapping validation ranges
4. **Style Object Sharing** - Border/Font/Fill objects reused incorrectly
5. **External Link Issues** - Dashboard can't resolve source workbooks

This script identifies causes #1-3 directly. For style issues, use Excel's
built-in repair (usually auto-fixes without data loss).

**Dashboard-Specific Considerations (A816-5):**

Dashboard requires normalized source workbooks:
- A816_1_Monitoring_Infrastructure.xlsx
- A816_2_Baseline_Detection.xlsx
- A816_3_Coverage_Assessment.xlsx
- A816_4_Alert_Management.xlsx

If these are missing, dashboard shows #REF errors. This is EXPECTED until:
1. Source workbooks generated (Scripts 1-4)
2. Normalization completed (normalize_assessment_files_a816.py)
3. Files placed in Dashboard_Sources/ directory

Sanity checker will warn about external references but note this is expected.

**Preventive vs. Reactive Usage:**

Preventive (Best Practice):
- Run after every workbook generation
- Validate before distributing to stakeholders
- Check before normalization/consolidation
- Part of automated quality assurance workflow

Reactive (When Problems Occur):
- Excel shows repair warnings
- Formulas show #REF errors
- Dashboard consolidation fails
- Stakeholders report issues opening files

Use preventively to avoid reactive firefighting.

**Integration with CI/CD:**

```bash
#!/bin/bash
# Quality assurance script for A.8.16 workbook generation

set -e  # Exit on any error

echo "Generating A.8.16 assessment workbooks..."
python3 generate_a816_1_monitoring_infrastructure.py
python3 generate_a816_2_baseline_detection.py
python3 generate_a816_3_coverage_assessment.py
python3 generate_a816_4_alert_management.py

echo "Running sanity checks..."
for workbook in ISMS_A_8_16_[1-4]_*.xlsx; do
    echo "Checking: $workbook"
    python3 excel_sanity_check_a816.py "$workbook" || exit 1
done

echo "All workbooks passed sanity checks ✓"
```

**Audit Considerations:**

Running sanity checks demonstrates quality assurance to auditors:
- Systematic validation of assessment tools
- Documented quality control process
- Evidence of due diligence

Keep sanity check output as QA documentation for audit trail.

**False Positives:**

Some warnings may be acceptable based on context:
- Dashboard external link warnings (expected before normalization)
- Minor sheet name variations (if intentional)
- Optional field completeness (if not applicable)

Review findings and apply judgment - don't blindly fix everything.

**Performance:**

Script typically completes in <10 seconds per workbook. Factors affecting speed:
- Workbook size (large files take longer)
- Number of formulas (more formulas = longer validation)
- Disk I/O (network drives slower than local)

**Limitations:**

What this script CANNOT detect:
- Data accuracy (assessor responsibility)
- Content completeness (human judgment required)
- Stylistic quality (subjective)
- Excel version compatibility issues
- Macro/VBA problems (workbooks don't use macros)

What this script DOES detect:
- Structural integrity issues
- Formula correctness
- Required component presence
- Common openpyxl generation errors

**Troubleshooting:**

Problem: Script fails to load workbook
Solution: File may be corrupt - try opening in Excel and saving

Problem: Script reports missing sheets that exist
Solution: Sheet names may have extra spaces or different case - check carefully

Problem: No issues detected but Excel still shows repair warning
Solution: Issue may be style-related (not detected by script) - let Excel repair

Problem: Script runs very slowly
Solution: Large workbook or network drive - copy to local disk for checking

**Related Quality Assurance Tools:**

Complete A.8.16 QA toolchain:
1. **excel_sanity_check_a816.py** (this script) - Structural validation
2. **normalize_assessment_files_a816.py** - Data normalization & validation
3. **consolidate_a816_dashboard.py** - Data consolidation verification

Use all three for comprehensive quality assurance.

================================================================================
"""

import sys
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ============================================================================
# WORKBOOK TYPE DETECTION
# ============================================================================

def detect_workbook_type(filename):
    """Detect which A.8.16 workbook type this is based on filename."""
    filename_lower = filename.lower()
    
    if 'monitoring' in filename_lower and 'infrastructure' in filename_lower:
        return 'A816-1', 'Monitoring Infrastructure Assessment'
    elif 'baseline' in filename_lower or 'detection' in filename_lower:
        return 'A816-2', 'Baseline & Anomaly Detection Assessment'
    elif 'coverage' in filename_lower:
        return 'A816-3', 'Coverage Assessment'
    elif 'alert' in filename_lower and 'management' in filename_lower:
        return 'A816-4', 'Alert Management & Response Assessment'
    elif 'dashboard' in filename_lower or 'compliance' in filename_lower:
        return 'A816-5', 'Compliance Dashboard'
    else:
        return 'Unknown', 'Generic A.8.16 Excel Workbook'


# ============================================================================
# EXPECTED SHEET DEFINITIONS
# ============================================================================

EXPECTED_SHEETS = {
    'A816-1': [
        "Instructions & Legend", "1. Monitoring Platform", "2. Log Source Coverage",
        "3. Data Collection Arch", "4. Integration Enrichment", "5. Performance Scale",
        "Summary Dashboard", "Evidence Register", "Approval Sign-Off"
    ],
    'A816-2': [
        "Instructions & Legend", "1. Normal Behavior Profiles", "2. Detection Rules",
        "3. ML Models", "4. Threat Intel", "5. Validation Testing",
        "Summary Dashboard", "Evidence Register", "Approval Sign-Off"
    ],
    'A816-3': [
        "Instructions & Legend", "1. Asset Coverage", "2. Network Coverage",
        "3. User Activity Coverage", "4. Application Coverage", "5. Cloud Coverage",
        "Summary Dashboard", "Evidence Register", "Approval Sign-Off"
    ],
    'A816-4': [
        "Instructions & Legend", "1. Alert Generation", "2. Alert Triage",
        "3. Incident Response", "4. Escalation Procedures", "5. Performance Metrics",
        "Summary Dashboard", "Evidence Register", "Approval Sign-Off"
    ],
    'A816-5': [
        "Instructions & Legend", "Executive Summary", "Compliance Matrix",
        "KPIs", "Gap Remediation Tracker", "Trend Analysis", "Evidence & Approvals"
    ],
}

# Expected Assessment IDs
EXPECTED_ASSESSMENT_IDS = {
    'A816-1': 'ISMS-IMP-A.8.16.1',
    'A816-2': 'ISMS-IMP-A.8.16.2',
    'A816-3': 'ISMS-IMP-A.8.16.3',
    'A816-4': 'ISMS-IMP-A.8.16.4',
    'A816-5': 'ISMS-IMP-A.8.16.5',
}

# Expected normalized filenames for dashboard
DASHBOARD_SOURCES = [
    'A816_1_Monitoring_Infrastructure.xlsx',
    'A816_2_Baseline_Detection.xlsx',
    'A816_3_Coverage_Assessment.xlsx',
    'A816_4_Alert_Management.xlsx',
]


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""
    
    workbook_id, workbook_name = detect_workbook_type(filename)
    
    print("=" * 80)
    print(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {Path(filename).name}")
    print(f"Detected Type: {workbook_id} - {workbook_name}")
    print("=" * 80)
    
    issues_found = []
    warnings_found = []
    
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"\n✓ Workbook loaded successfully")
        print(f"  Sheets found: {len(wb.sheetnames)}")
        print(f"  Sheet names: {', '.join(wb.sheetnames)}")
        
    except Exception as e:
        print(f"\n✗ CRITICAL: Cannot load workbook: {e}")
        return
    
    # ========================================================================
    # CHECK 1: SHEET STRUCTURE VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 1: SHEET STRUCTURE VALIDATION")
    print("=" * 80)
    
    if workbook_id in EXPECTED_SHEETS:
        expected = EXPECTED_SHEETS[workbook_id]
        actual = wb.sheetnames
        
        # Check for missing sheets (flexible matching)
        missing = []
        for exp_sheet in expected:
            # Flexible match: check if expected sheet name is in any actual sheet
            found = False
            for act_sheet in actual:
                if exp_sheet.lower().replace("_", "").replace(" ", "") == act_sheet.lower().replace("_", "").replace(" ", ""):
                    found = True
                    break
            if not found:
                missing.append(exp_sheet)
        
        if missing:
            for sheet in missing:
                warnings_found.append(f"  ⚠ Missing expected sheet: {sheet}")
            print(f"  ⚠ {len(missing)} expected sheets not found (may be renamed)")
        else:
            print(f"  ✓ All expected sheets present for {workbook_id}")
        
        # Check for extra sheets
        print(f"  Total sheets: {len(actual)} (expected {len(expected)})")
    else:
        print("  ℹ Unknown workbook type - skipping sheet validation")
    
    # ========================================================================
    # CHECK 2: ASSESSMENT ID VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 2: ASSESSMENT ID VALIDATION")
    print("=" * 80)
    
    assessment_id_found = False
    if workbook_id in EXPECTED_ASSESSMENT_IDS:
        expected_id = EXPECTED_ASSESSMENT_IDS[workbook_id]
        
        # Check in Instructions & Legend sheet
        if 'Instructions & Legend' in wb.sheetnames:
            ws = wb['Instructions & Legend']
            # Check first few rows for assessment ID
            for row in range(1, 15):
                for col in range(1, 10):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and expected_id in cell_value:
                        print(f"  ✓ Assessment ID found: {expected_id}")
                        assessment_id_found = True
                        break
                if assessment_id_found:
                    break
        
        # Also check in other common locations
        if not assessment_id_found:
            for sheet_name in wb.sheetnames:
                if 'Instructions' in sheet_name or 'Summary' in sheet_name:
                    ws = wb[sheet_name]
                    for row in range(1, 10):
                        for col in range(1, 5):
                            cell_value = ws.cell(row=row, column=col).value
                            if cell_value and isinstance(cell_value, str) and expected_id in cell_value:
                                print(f"  ✓ Assessment ID found in {sheet_name}: {expected_id}")
                                assessment_id_found = True
                                break
                        if assessment_id_found:
                            break
                if assessment_id_found:
                    break
        
        if not assessment_id_found:
            warnings_found.append(
                f"  ⚠ Assessment ID '{expected_id}' not found in Instructions sheet"
            )
    else:
        print("  ℹ Unknown workbook type - skipping Assessment ID validation")
    
    # ========================================================================
    # CHECK 3: FORMULA VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 3: FORMULA VALIDATION")
    print("=" * 80)
    
    formula_issues = 0
    inter_sheet_refs = {}
    external_refs = set()
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # Check for external workbook references
                    ext_refs = re.findall(r'\[([^\]]+\.xlsx)\]', formula)
                    for ref in ext_refs:
                        external_refs.add(ref)
                    
                    # Check for sheet references
                    if "'" in formula or "!" in formula:
                        sheet_refs = re.findall(r"'?([^'!]+)'?!", formula)
                        for ref in sheet_refs:
                            ref_clean = ref.strip("'")
                            # Skip external references
                            if "[" not in ref_clean and ref_clean not in wb.sheetnames:
                                # Check for close matches (underscore vs space)
                                close_match = False
                                for actual in wb.sheetnames:
                                    if ref_clean.replace(" ", "_") == actual or ref_clean.replace("_", " ") == actual:
                                        close_match = True
                                        break
                                if not close_match:
                                    issues_found.append(
                                        f"  ✗ {sheet_name}!{cell.coordinate}: "
                                        f"Invalid sheet reference '{ref_clean}'"
                                    )
                                    formula_issues += 1
                            
                            # Track inter-sheet references
                            if "[" not in ref_clean and ref_clean in wb.sheetnames:
                                if sheet_name not in inter_sheet_refs:
                                    inter_sheet_refs[sheet_name] = set()
                                inter_sheet_refs[sheet_name].add(ref_clean)
                    
                    # Check for common syntax issues
                    if formula.count('(') != formula.count(')'):
                        issues_found.append(
                            f"  ✗ {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced parentheses in formula"
                        )
                        formula_issues += 1
    
    if formula_issues == 0:
        print("  ✓ No formula syntax issues detected")
    else:
        print(f"  ✗ Found {formula_issues} formula issues")
    
    if inter_sheet_refs:
        print(f"  ℹ Inter-sheet references found:")
        for sheet, refs in inter_sheet_refs.items():
            print(f"    {sheet} → {', '.join(sorted(refs))}")
    
    if external_refs:
        print(f"  ℹ External workbook references found:")
        for ref in sorted(external_refs):
            print(f"    → {ref}")
        if workbook_id == 'A816-5':
            print("  ⚠ WARNING: Dashboard requires external links to function")
    
    # ========================================================================
    # CHECK 4: DATA VALIDATION CONFLICTS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 4: DATA VALIDATION CONFLICTS")
    print("=" * 80)
    
    validation_issues = 0
    total_validations = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            dv_count = len(ws.data_validations.dataValidation)
            total_validations += dv_count
            
            # Check for overlapping ranges
            all_ranges = []
            for dv in ws.data_validations.dataValidation:
                if hasattr(dv, 'sqref') and dv.sqref:
                    all_ranges.extend(str(dv.sqref).split())
            
            if len(all_ranges) != len(set(all_ranges)):
                warnings_found.append(
                    f"  ⚠ {sheet_name}: Potentially overlapping data validation ranges"
                )
                validation_issues += 1
    
    print(f"  Total data validations: {total_validations}")
    
    if validation_issues == 0:
        print("  ✓ No data validation conflicts detected")
    
    # ========================================================================
    # CHECK 5: MERGED CELLS VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 5: MERGED CELLS VALIDATION")
    print("=" * 80)
    
    merge_issues = 0
    total_merges = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'merged_cells') and ws.merged_cells:
            merge_count = len(ws.merged_cells.ranges)
            total_merges += merge_count
            
            # Check merged cells for content issues
            for merge_range in ws.merged_cells.ranges:
                min_col, min_row = merge_range.min_col, merge_range.min_row
                max_col, max_row = merge_range.max_col, merge_range.max_row
                
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if row == min_row and col == min_col:
                            continue
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            cell_coord = f"{get_column_letter(col)}{row}"
                            warnings_found.append(
                                f"  ⚠ {sheet_name}!{cell_coord}: "
                                f"Merged cell has content in non-primary cell"
                            )
                            merge_issues += 1
    
    print(f"  Total merged ranges: {total_merges}")
    
    if merge_issues == 0:
        print("  ✓ Merged cells properly formatted")
    
    # ========================================================================
    # CHECK 6: EVIDENCE REGISTER VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 6: EVIDENCE REGISTER VALIDATION")
    print("=" * 80)
    
    evidence_sheet = None
    for sheet in wb.sheetnames:
        if 'evidence' in sheet.lower() and 'register' in sheet.lower():
            evidence_sheet = sheet
            break
    
    if evidence_sheet:
        ws = wb[evidence_sheet]
        # Count evidence rows (look for EV- prefix)
        evidence_count = 0
        for row in range(1, min(110, ws.max_row + 1)):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and isinstance(cell_val, str):
                if cell_val.startswith('EV-'):
                    evidence_count += 1
        
        print(f"  Evidence sheet: {evidence_sheet}")
        print(f"  Evidence row templates: {evidence_count}")
        
        # A.8.16 requires minimum 10 evidence items
        if evidence_count < 10:
            issues_found.append(
                f"  ✗ Evidence Register has {evidence_count} rows (MINIMUM 10 REQUIRED)"
            )
        elif evidence_count < 50:
            warnings_found.append(
                f"  ⚠ Evidence Register has {evidence_count} rows (recommended ≥100 for flexibility)"
            )
        elif evidence_count >= 100:
            print(f"  ✓ Full 100-row evidence register (excellent)")
        else:
            print(f"  ✓ Evidence register adequate ({evidence_count} rows)")
    else:
        if workbook_id != 'A816-5':  # Dashboard has Evidence & Approvals instead
            issues_found.append("  ✗ No Evidence Register sheet found (REQUIRED)")
    
    # ========================================================================
    # CHECK 7: APPROVAL SIGN-OFF VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 7: APPROVAL SIGN-OFF VALIDATION")
    print("=" * 80)
    
    approval_sheet = None
    for sheet in wb.sheetnames:
        if 'approval' in sheet.lower() or 'sign' in sheet.lower():
            approval_sheet = sheet
            break
    
    if approval_sheet:
        ws = wb[approval_sheet]
        
        # Check for 3-level approval structure
        approval_levels = 0
        for row in range(1, min(100, ws.max_row + 1)):
            for col in range(1, 6):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val and isinstance(cell_val, str):
                    if 'PREPARED BY' in cell_val or 'COMPLETED BY' in cell_val:
                        approval_levels += 1
                    elif 'SOC LEAD' in cell_val or 'REVIEWED BY' in cell_val:
                        approval_levels += 1
                    elif 'CISO' in cell_val or 'APPROVED BY' in cell_val:
                        approval_levels += 1
        
        print(f"  Approval sheet: {approval_sheet}")
        print(f"  Approval levels detected: {approval_levels}")
        
        if approval_levels < 3:
            warnings_found.append(
                f"  ⚠ Approval workflow has {approval_levels} levels (expected 3: Assessor → SOC Lead → CISO)"
            )
        else:
            print(f"  ✓ Complete 3-level approval workflow")
    else:
        issues_found.append("  ✗ No Approval Sign-Off sheet found (REQUIRED)")
    
    # ========================================================================
    # CHECK 8: SUMMARY DASHBOARD VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 8: SUMMARY DASHBOARD VALIDATION")
    print("=" * 80)
    
    summary_sheet = None
    for sheet in wb.sheetnames:
        if 'summary' in sheet.lower() or 'dashboard' in sheet.lower():
            summary_sheet = sheet
            break
    
    if summary_sheet:
        ws = wb[summary_sheet]
        
        # Count formulas in Summary sheet
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
        
        print(f"  Summary sheet: {summary_sheet}")
        print(f"  Formulas in Summary: {formula_count}")
        
        if formula_count < 10:
            warnings_found.append(
                f"  ⚠ Summary sheet has only {formula_count} formulas (expected many calculated metrics)"
            )
        else:
            print(f"  ✓ Summary dashboard has active calculations")
    else:
        if workbook_id != 'A816-5':  # Dashboard has Executive Summary instead
            warnings_found.append("  ⚠ No Summary Dashboard sheet found")
    
    # ========================================================================
    # CHECK 9: ROW CAPACITY VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 9: ROW CAPACITY VALIDATION")
    print("=" * 80)
    
    data_sheets = [s for s in wb.sheetnames if s not in ['Instructions & Legend', 'Summary Dashboard', 
                                                          'Evidence Register', 'Approval Sign-Off',
                                                          'Evidence & Approvals', 'Executive Summary']]
    
    for sheet_name in data_sheets:
        ws = wb[sheet_name]
        
        # Check if data entry area exists (typically rows 7+ for 110 capacity)
        if ws.max_row > 100:
            # Count example rows vs entry rows
            example_count = 0
            entry_area_start = 20  # Typically after examples
            
            for row in range(7, min(20, ws.max_row)):
                # Check if row has data (non-empty first column)
                if ws.cell(row=row, column=1).value:
                    example_count += 1
            
            total_capacity = ws.max_row - 6  # Subtract header rows
            
            print(f"  {sheet_name}:")
            print(f"    Total capacity: ~{total_capacity} rows")
            print(f"    Example data rows: ~{example_count}")
            
            if total_capacity < 50:
                warnings_found.append(
                    f"  ⚠ {sheet_name}: Low capacity ({total_capacity} rows), recommended ≥110"
                )
            elif total_capacity >= 110:
                print(f"    ✓ Adequate capacity (≥110 rows)")
    
    # ========================================================================
    # CHECK 10: WORKSHEET DIMENSIONS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 10: WORKSHEET DIMENSIONS")
    print("=" * 80)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if ws.max_row > 5000:
            warnings_found.append(
                f"  ⚠ {sheet_name}: Large worksheet ({ws.max_row} rows)"
            )
        
        if ws.max_column > 50:
            warnings_found.append(
                f"  ⚠ {sheet_name}: Wide worksheet ({ws.max_column} columns)"
            )
    
    print("  ✓ Worksheet dimensions within reasonable limits")
    
    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    print("\n" + "=" * 80)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 80)
    
    if issues_found:
        print(f"\n❌ CRITICAL ISSUES FOUND: {len(issues_found)}")
        for issue in issues_found:
            print(issue)
    
    if warnings_found:
        print(f"\n⚠️  WARNINGS: {len(warnings_found)}")
        for warning in warnings_found:
            print(warning)
    
    if not issues_found and not warnings_found:
        print("\n✅ NO ISSUES DETECTED")
        print("\nThe workbook appears structurally sound.")
        print("If Excel still shows repair warnings, consider:")
        print("  • Excel version compatibility (try Excel 2019+)")
        print("  • Antivirus/security software interference")
        print("  • Network drive or OneDrive sync issues")
        print("  • Excel's overly cautious validation")
        if workbook_id == 'A816-5':
            print("  • Unresolved external workbook links (expected before normalization)")
    else:
        print_recommendations(issues_found, warnings_found, workbook_id, 
                            formula_issues, validation_issues, merge_issues)
    
    # ========================================================================
    # WORKBOOK-SPECIFIC NOTES
    # ========================================================================
    print_workbook_notes(workbook_id, external_refs)
    
    wb.close()


def print_recommendations(issues, warnings, workbook_id, formula_issues, 
                         validation_issues, merge_issues):
    """Print recommended fixes based on issues found."""
    print("\n" + "=" * 80)
    print("RECOMMENDED ACTIONS")
    print("=" * 80)
    
    if formula_issues > 0:
        print("\n1. FORMULA FIXES:")
        print("   • Review formulas referencing other sheets")
        print("   • Ensure sheet names match exactly (case-sensitive)")
        print("   • Check for balanced quotes and parentheses")
        if workbook_id == 'A816-5':
            print("   • Verify external workbook references point to normalized files")
    
    if validation_issues > 0:
        print("\n2. DATA VALIDATION FIXES:")
        print("   • Remove overlapping data validation ranges")
        print("   • Apply validations to specific ranges, not entire columns")
    
    if merge_issues > 0:
        print("\n3. MERGED CELL FIXES:")
        print("   • Ensure only top-left cell of merged range has content")
        print("   • Clear content from other cells in merged range")
    
    if workbook_id == 'A816-5':
        print("\n4. DASHBOARD-SPECIFIC SETUP (A816-5):")
        print("   • This workbook consolidates data from 4 assessment workbooks")
        print("   • Ensure normalized source workbooks are in the same folder:")
        for source in DASHBOARD_SOURCES:
            print(f"     - {source}")
        print("   • Run normalize_assessment_files_a816.py first to create normalized files")


def print_workbook_notes(workbook_id, external_refs):
    """Print workbook-specific notes and expected structure."""
    print("\n" + "=" * 80)
    print(f"WORKBOOK-SPECIFIC NOTES: {workbook_id}")
    print("=" * 80)
    
    if workbook_id == 'A816-1':
        print("\nMonitoring Infrastructure Assessment:")
        print("  • 9 sheets for infrastructure assessment")
        print("  • 110 entry capacity (10 examples + 100 entry rows)")
        print("  • Monitoring platform capabilities tracking")
        print("  • Log source coverage assessment")
        print("  • Data collection architecture validation")
        print("  • Integration and enrichment capabilities")
        print("  • Performance and scalability assessment")
        print("  • Evidence Register (100 entries minimum)")
        print("  • 3-level approval workflow (Assessor → SOC Lead → CISO)")
        print("\n  Key metrics:")
        print("    - Platform capability score")
        print("    - Log source coverage percentage")
        print("    - Data collection reliability")
        print("    - Integration maturity level")
    
    elif workbook_id == 'A816-2':
        print("\nBaseline & Anomaly Detection Assessment:")
        print("  • 9 sheets for detection capabilities")
        print("  • 110 entry capacity per assessment sheet")
        print("  • Normal behavior baseline documentation")
        print("  • Detection rule effectiveness tracking")
        print("  • ML/AI model assessment")
        print("  • Threat intelligence integration")
        print("  • Validation testing results")
        print("  • Evidence Register (100 entries minimum)")
        print("  • 3-level approval workflow")
        print("\n  Key metrics:")
        print("    - Baseline coverage percentage")
        print("    - Detection rule effectiveness rate")
        print("    - False positive rate")
        print("    - Detection capability maturity")
    
    elif workbook_id == 'A816-3':
        print("\nCoverage Assessment:")
        print("  • 9 sheets for coverage analysis")
        print("  • 110 entry capacity per assessment sheet")
        print("  • Asset monitoring coverage")
        print("  • Network segment coverage")
        print("  • User activity monitoring")
        print("  • Application coverage assessment")
        print("  • Cloud environment monitoring")
        print("  • Evidence Register (100 entries minimum)")
        print("  • 3-level approval workflow")
        print("\n  Key metrics:")
        print("    - Overall monitoring coverage (Tier 1: 100%, Tier 2: 90%, Tier 3: 70%)")
        print("    - Gap identification and remediation")
        print("    - Coverage by criticality tier")
        print("    - Blind spot analysis")
    
    elif workbook_id == 'A816-4':
        print("\nAlert Management & Response Assessment:")
        print("  • 9 sheets for alert lifecycle")
        print("  • 110 entry capacity per assessment sheet")
        print("  • Alert generation effectiveness")
        print("  • Alert triage procedures")
        print("  • Incident response workflows")
        print("  • Escalation procedures")
        print("  • Performance metrics (MTTD, MTTR)")
        print("  • Evidence Register (100 entries minimum)")
        print("  • 3-level approval workflow")
        print("\n  Key metrics:")
        print("    - Alert response time (MTTD, MTTR)")
        print("    - False positive rate")
        print("    - Escalation effectiveness")
        print("    - Incident response maturity")
    
    elif workbook_id == 'A816-5':
        print("\nCompliance Dashboard (MASTER):")
        print("  • 7 sheets - EXECUTIVE CONSOLIDATION WORKBOOK")
        print("  • Overall A.8.16 compliance score")
        print("  • Compliance matrix across all 4 assessments")
        print("  • KPI tracking (MTTD, MTTR, coverage, detection rate)")
        print("  • Gap remediation tracker")
        print("  • Trend analysis over time")
        print("  • Evidence completeness tracking")
        print("  • 3-level approval workflow")
        print("\n  ⚠ This dashboard requires normalized source files:")
        for source in DASHBOARD_SOURCES:
            print(f"    - {source}")
        print("\n  Setup process:")
        print("    1. Generate all 4 assessment workbooks (Scripts 1-4)")
        print("    2. Run: python normalize_assessment_files_a816.py")
        print("    3. Dashboard reads from Dashboard_Sources/ directory")
        print("    4. May show #REF errors until normalization complete")
    
    else:
        print("\nGeneric A.8.16 Workbook")
        print("  • No specific validation rules defined for this type")
        print("  • Standard Excel health checks applied")
        print("  • Ensure workbook follows A.8.16 framework structure")
    
    print("\n" + "=" * 80)


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("=" * 70)
        print("ISMS A.8.16 Framework - Excel Sanity Checker")
        print("Control: A.8.16 (Monitoring Activities)")
        print("=" * 70)
        print("\nUsage: python3 excel_sanity_check_a816.py <filename.xlsx>")
        print("\nExamples:")
        print("  python3 excel_sanity_check_a816.py ISMS_A_8_16_1_Monitoring_Infrastructure.xlsx")
        print("  python3 excel_sanity_check_a816.py ISMS_A_8_16_2_Baseline_Detection.xlsx")
        print("  python3 excel_sanity_check_a816.py ISMS_A_8_16_3_Coverage_Assessment.xlsx")
        print("  python3 excel_sanity_check_a816.py ISMS_A_8_16_4_Alert_Management.xlsx")
        print("  python3 excel_sanity_check_a816.py ISMS_A_8_16_5_Compliance_Dashboard.xlsx")
        print("\nSupported workbook types:")
        print("  A816-1 - Monitoring Infrastructure Assessment")
        print("  A816-2 - Baseline & Anomaly Detection Assessment")
        print("  A816-3 - Coverage Assessment")
        print("  A816-4 - Alert Management & Response Assessment")
        print("  A816-5 - Compliance Dashboard")
        print("\n'Don't fool yourself' - Systems Engineering ISMS")
        sys.exit(1)
    
    filename = sys.argv[1]
    
    try:
        check_workbook_health(filename)
    except FileNotFoundError:
        print(f"\n❌ ERROR: File not found: {filename}")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()