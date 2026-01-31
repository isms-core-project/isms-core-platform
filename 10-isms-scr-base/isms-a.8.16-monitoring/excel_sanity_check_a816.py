#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
Excel Workbook Sanity Checker for ISMS Control A.8.16 Monitoring Activities Workbooks
Diagnoses issues that trigger Excel's "file level validation and repair"

GENERIC VERSION - Works for all A.8.16 assessment workbooks (A.8.16.1 through A.8.16.5)

Usage:
    python3 excel_sanity_check_a816.py ISMS-IMP-A.8.16.1_Monitoring_Infrastructure_20260107.xlsx
    python3 excel_sanity_check_a816.py ISMS-IMP-A.8.16.5_Compliance_Dashboard_20260107.xlsx

Philosophy:
    "You must not fool yourself — and you are the easiest person to fool."
    - Richard Feynman
    
    This script detects cargo cult Excel issues before your auditor does.

"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import re
import sys

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
def detect_workbook_type(filename):
    """Detect which A.8.16 workbook type this is based on filename."""
    filename_lower = filename.lower()
    
    if 'a.8.16.1' in filename_lower or 'monitoring_infrastructure' in filename_lower:
        return 'A.8.16.1', 'Monitoring Infrastructure Assessment'
    elif 'a.8.16.2' in filename_lower or 'baseline_detection' in filename_lower or 'baseline' in filename_lower:
        return 'A.8.16.2', 'Baseline & Detection Assessment'
    elif 'a.8.16.3' in filename_lower or 'coverage_assessment' in filename_lower or 'coverage' in filename_lower:
        return 'A.8.16.3', 'Coverage Assessment'
    elif 'a.8.16.4' in filename_lower or 'alert_management' in filename_lower or 'alert' in filename_lower:
        return 'A.8.16.4', 'Alert Management & Response Assessment'
    elif 'a.8.16.5' in filename_lower or 'compliance_dashboard' in filename_lower:
        return 'A.8.16.5', 'Compliance Dashboard'
    else:
        return 'Unknown', 'Generic Excel Workbook'


# ============================================================================
# EXPECTED SHEET DEFINITIONS
# ============================================================================

EXPECTED_SHEETS = {
    'A.8.16.1': [
        "Instructions & Legend", "1. Monitoring Platform", "2. Log Source Coverage",
        "3. Data Collection Arch", "4. Integration Enrichment", "5. Performance Scale",
        "Summary Dashboard", "Evidence Register", "Approval Sign-Off"
    ],
    'A.8.16.2': [
        "Instructions & Legend", "1. Baseline Inventory", "2. Detection Rules",
        "3. MITRE ATT&CK Coverage", "4. Rule Performance", "5. Testing Validation",
        "Summary Dashboard", "Evidence Register", "Approval Sign-Off"
    ],
    'A.8.16.3': [
        "Instructions & Legend", "1. Asset Coverage", "2. Network Coverage",
        "3. User Identity Coverage", "4. Application Coverage", "5. Gap Analysis",
        "Summary Dashboard", "Evidence Register", "Approval Sign-Off"
    ],
    'A.8.16.4': [
        "Instructions & Legend", "1. Alert Generation", "2. Triage Investigation",
        "3. Escalation Response", "4. Performance Metrics", "5. SOC Readiness",
        "Summary Dashboard", "Evidence Register", "Approval Sign-Off"
    ],
    'A.8.16.5': [
        "Instructions & Legend", "Executive Summary", "Compliance Matrix",
        "KPIs", "Gap Remediation Tracker", "Trend Analysis", "Evidence & Approvals"
    ],
}


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""
    
    workbook_id, workbook_name = detect_workbook_type(filename)
    
    print("=" * 80)
    print(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {filename}")
    print(f"Detected Type: {workbook_id} - {workbook_name}")
    print("=" * 80)
    
    issues_found = []
    warnings_found = []
    
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"\n✓ Workbook loaded successfully")
        print(f"  Sheets found: {len(wb.sheetnames)}")
        print(f"  Sheet names: {', '.join(wb.sheetnames)}")
        
    except Exception as e:
        print(f"\n✗ CRITICAL: Cannot load workbook: {e}")
        return
    
    # ========================================================================
    # CHECK 1: SHEET STRUCTURE VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 1: SHEET STRUCTURE VALIDATION")
    print("=" * 80)
    
    if workbook_id in EXPECTED_SHEETS:
        expected = EXPECTED_SHEETS[workbook_id]
        actual = wb.sheetnames
        
        # Check for missing sheets (flexible matching)
        missing = []
        for exp_sheet in expected:
            # Flexible match: check if expected sheet name is in any actual sheet
            found = False
            for act_sheet in actual:
                if exp_sheet.lower().replace("_", "").replace("&", "").replace(" ", "") in \
                   act_sheet.lower().replace("_", "").replace("&", "").replace(" ", ""):
                    found = True
                    break
            if not found:
                missing.append(exp_sheet)
        
        if missing:
            for sheet in missing:
                warnings_found.append(f"  ⚠ Missing expected sheet: {sheet}")
            print(f"  ⚠ {len(missing)} expected sheets not found (may be renamed)")
        else:
            print(f"  ✓ All expected sheets present for {workbook_id}")
        
        # Check for extra sheets
        print(f"  Total sheets: {len(actual)} (expected {len(expected)})")
    else:
        print("  ℹ Unknown workbook type - skipping sheet validation")
    
    # ========================================================================
    # CHECK 2: FORMULA VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 2: FORMULA VALIDATION")
    print("=" * 80)
    
    formula_issues = 0
    inter_sheet_refs = {}
    external_refs = set()
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # Check for external workbook references
                    ext_refs = re.findall(r'\[([^\]]+\.xlsx)\]', formula)
                    for ref in ext_refs:
                        external_refs.add(ref)
                    
                    # Check for sheet references
                    if "'" in formula or "!" in formula:
                        sheet_refs = re.findall(r"'?([^'!]+)'?!", formula)
                        for ref in sheet_refs:
                            ref_clean = ref.strip("'")
                            # Skip external references
                            if "[" not in ref_clean and ref_clean not in wb.sheetnames:
                                # Check for close matches (underscore vs space vs &)
                                close_match = False
                                for actual in wb.sheetnames:
                                    if ref_clean.replace(" ", "_").replace("&", "") == \
                                       actual.replace(" ", "_").replace("&", "") or \
                                       ref_clean.replace("_", " ").replace("&", "") == \
                                       actual.replace("_", " ").replace("&", ""):
                                        close_match = True
                                        break
                                if not close_match:
                                    issues_found.append(
                                        f"  ✗ {sheet_name}!{cell.coordinate}: "
                                        f"Invalid sheet reference '{ref_clean}'"
                                    )
                                    formula_issues += 1
                            elif ref_clean in wb.sheetnames:
                                # Track inter-sheet references
                                if sheet_name not in inter_sheet_refs:
                                    inter_sheet_refs[sheet_name] = set()
                                inter_sheet_refs[sheet_name].add(ref_clean)
    
    if formula_issues == 0:
        print(f"  ✓ No formula issues detected")
    else:
        print(f"  ✗ {formula_issues} formula issues found")
    
    if inter_sheet_refs:
        print(f"\n  Inter-sheet references found (expected):")
        for source, targets in sorted(inter_sheet_refs.items()):
            print(f"    {source} → {', '.join(sorted(targets))}")
    
    if external_refs:
        print(f"\n  External workbook references found:")
        for ref in sorted(external_refs):
            print(f"    [{ref}]")
        if workbook_id == 'A.8.16.5':
            print(f"  ℹ External references are EXPECTED in the Compliance Dashboard")
        else:
            warnings_found.append(f"  ⚠ External references found in non-dashboard workbook")
    
    # ========================================================================
    # CHECK 3: DATA VALIDATION RULES
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 3: DATA VALIDATION RULES")
    print("=" * 80)
    
    validation_issues = 0
    validation_overlaps = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Check for overlapping validation ranges
        if ws.data_validations and len(ws.data_validations.dataValidation) > 0:
            ranges = []
            for dv in ws.data_validations.dataValidation:
                if hasattr(dv, 'sqref') and dv.sqref:
                    ranges.append((dv.sqref, dv))
            
            # Simple overlap detection: if same column appears in multiple rules
            cols_seen = {}
            for sqref, dv in ranges:
                for cell_range in str(sqref).split():
                    col = ''.join(filter(str.isalpha, cell_range.split(':')[0]))
                    if col in cols_seen:
                        validation_overlaps.append(
                            f"  ⚠ {sheet_name}: Column {col} has multiple validations"
                        )
                        validation_issues += 1
                    cols_seen[col] = True
    
    if validation_issues == 0:
        print(f"  ✓ No data validation issues detected")
    else:
        print(f"  ⚠ {validation_issues} potential validation overlaps found")
        for overlap in validation_overlaps[:10]:  # Limit to first 10
            warnings_found.append(overlap)
    
    # ========================================================================
    # CHECK 4: MERGED CELLS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 4: MERGED CELLS VALIDATION")
    print("=" * 80)
    
    merge_issues = 0
    merge_problems = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for merged_range in ws.merged_cells.ranges:
            # Get the top-left cell of merged range
            top_left = merged_range.min_row, merged_range.min_col
            
            # Check if any non-top-left cell in the merged range has content
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if (row, col) != top_left:
                        cell = ws.cell(row, col)
                        if cell.value:
                            merge_problems.append(
                                f"  ✗ {sheet_name}!{cell.coordinate}: "
                                f"Merged cell has content in non-top-left position"
                            )
                            merge_issues += 1
    
    if merge_issues == 0:
        print(f"  ✓ Merged cells properly configured")
    else:
        print(f"  ✗ {merge_issues} merged cell issues found")
        for problem in merge_problems[:10]:  # Limit to first 10
            issues_found.append(problem)
    
    # ========================================================================
    # CHECK 5: COLUMN WIDTHS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 5: COLUMN WIDTHS")
    print("=" * 80)
    
    width_warnings = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col_letter in ws.column_dimensions:
            width = ws.column_dimensions[col_letter].width
            if width and width > 255:
                warnings_found.append(
                    f"  ⚠ {sheet_name}: Column {col_letter} width {width} exceeds Excel max (255)"
                )
                width_warnings += 1
    
    if width_warnings == 0:
        print(f"  ✓ All column widths within valid range")
    else:
        print(f"  ⚠ {width_warnings} columns exceed maximum width")
    
    # ========================================================================
    # CHECK 6: PROTECTION & SECURITY
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 6: SHEET PROTECTION")
    print("=" * 80)
    
    protected_sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.protection.sheet:
            protected_sheets.append(sheet_name)
    
    if protected_sheets:
        print(f"  ℹ {len(protected_sheets)} sheets are protected:")
        for sheet in protected_sheets[:5]:  # Show first 5
            print(f"    - {sheet}")
    else:
        print(f"  ℹ No sheet protection enabled")
    
    # ========================================================================
    # DIAGNOSTIC SUMMARY
    # ========================================================================
    print("\n" + "=" * 80)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 80)
    
    if issues_found:
        print(f"\n❌ CRITICAL ISSUES FOUND: {len(issues_found)}")
        for issue in issues_found:
            print(issue)
    
    if warnings_found:
        print(f"\n⚠️ WARNINGS: {len(warnings_found)}")
        for warning in warnings_found:
            print(warning)
    
    if not issues_found and not warnings_found:
        print("\n✅ NO ISSUES DETECTED")
        print("\nThe workbook appears structurally sound.")
        print("If Excel still shows repair warnings, consider:")
        print("  • Excel version compatibility (try Excel 2019+)")
        print("  • Antivirus/security software interference")
        print("  • Network drive or OneDrive sync issues")
        print("  • Excel's overly cautious validation (false positive)")
        if workbook_id == 'A.8.16.5':
            print("  • Unresolved external workbook links (expected before setup)")
    else:
        print_recommendations(issues_found, warnings_found, workbook_id, 
                            formula_issues, validation_issues, merge_issues)
    
    # ========================================================================
    # WORKBOOK-SPECIFIC NOTES
    # ========================================================================
    print_workbook_notes(workbook_id, external_refs)
    
    wb.close()


def print_recommendations(issues, warnings, workbook_id, formula_issues, 
                         validation_issues, merge_issues):
    """Print recommended fixes based on issues found."""
    print("\n" + "=" * 80)
    print("RECOMMENDED ACTIONS")
    print("=" * 80)
    
    if formula_issues > 0:
        print("\n1. FORMULA FIXES:")
        print("   • Review formulas referencing other sheets")
        print("   • Ensure sheet names match exactly (case-sensitive)")
        print("   • Check for balanced quotes and parentheses")
        print("   • Watch for special characters in sheet names (& symbol)")
        if workbook_id == 'A.8.16.5':
            print("   • Verify external workbook references are correct")
    
    if validation_issues > 0:
        print("\n2. DATA VALIDATION FIXES:")
        print("   • Remove overlapping data validation ranges")
        print("   • Apply validations to specific ranges, not entire columns")
        print("   • Ensure dropdown lists reference valid ranges")
    
    if merge_issues > 0:
        print("\n3. MERGED CELL FIXES:")
        print("   • Ensure only top-left cell of merged range has content")
        print("   • Clear content from other cells in merged range")
        print("   • Re-apply merge after clearing conflicting cells")
    
    if workbook_id == 'A.8.16.5':
        print("\n4. DASHBOARD-SPECIFIC SETUP (A.8.16.5):")
        print("   • This workbook consolidates data from Assessments 1-4")
        print("   • Ensure source workbooks are in the same folder:")
        print("     - ISMS-IMP-A.8.16.1_Monitoring_Infrastructure_YYYYMMDD.xlsx")
        print("     - ISMS-IMP-A.8.16.2_Baseline_Detection_YYYYMMDD.xlsx")
        print("     - ISMS-IMP-A.8.16.3_Coverage_Assessment_YYYYMMDD.xlsx")
        print("     - ISMS-IMP-A.8.16.4_Alert_Management_YYYYMMDD.xlsx")
        print("   • Update external links in Excel: Data > Edit Links")


def print_workbook_notes(workbook_id, external_refs):
    """Print workbook-specific notes and expected structure."""
    print("\n" + "=" * 80)
    print(f"WORKBOOK-SPECIFIC NOTES: {workbook_id}")
    print("=" * 80)
    
    if workbook_id == 'A.8.16.1':
        print("\nMonitoring Infrastructure Assessment:")
        print("  • 9 sheets for platform and architecture assessment")
        print("  • Monitoring platform capabilities (20 entries)")
        print("  • Log source coverage inventory (50 sources)")
        print("  • Data collection architecture documentation")
        print("  • Integration and enrichment assessment")
        print("  • Performance and scalability baseline")
        print("  • Summary dashboard with compliance scoring")
        print("  • Evidence register (50 entries)")
        print("  • 3-level approval workflow")
        print("\n  Focus: Platform capabilities, not just installed tools")
    
    elif workbook_id == 'A.8.16.2':
        print("\nBaseline & Detection Assessment:")
        print("  • 9 sheets for baseline and detection rule evaluation")
        print("  • Baseline inventory (30 baselines)")
        print("  • Detection rules catalog (100 rules)")
        print("  • MITRE ATT&CK coverage mapping")
        print("  • Rule performance metrics")
        print("  • Testing and validation results")
        print("  • Summary dashboard with coverage analysis")
        print("  • Evidence register (50 entries)")
        print("  • Approval sign-off")
        print("\n  Philosophy: Detection that actually detects, not just logs")
    
    elif workbook_id == 'A.8.16.3':
        print("\nCoverage Assessment:")
        print("  • 9 sheets for comprehensive coverage analysis")
        print("  • Asset coverage inventory (100 assets)")
        print("  • Network segment coverage (30 segments)")
        print("  • User/identity coverage (50 identities)")
        print("  • Application coverage (40 applications)")
        print("  • Gap analysis with risk assessment")
        print("  • Summary dashboard with coverage matrices")
        print("  • Evidence register (50 entries)")
        print("  • Approval workflow")
        print("\n  Principle: Monitor what matters, not just what's easy")
    
    elif workbook_id == 'A.8.16.4':
        print("\nAlert Management & Response Assessment:")
        print("  • 9 sheets for alert lifecycle and SOC operations")
        print("  • Alert generation inventory (50 alerts)")
        print("  • Triage and investigation procedures")
        print("  • Escalation and response workflows")
        print("  • Performance metrics and SLAs")
        print("  • SOC readiness assessment")
        print("  • Summary dashboard with operational KPIs")
        print("  • Evidence register (50 entries)")
        print("  • Approval sign-off")
        print("\n  Reality check: Alerts are only useful if someone responds")
    
    elif workbook_id == 'A.8.16.5':
        print("\nCompliance Dashboard (CONSOLIDATION):")
        print("  • 7 sheets - MASTER CONSOLIDATION WORKBOOK")
        print("  • Executive summary with compliance scorecard")
        print("  • Compliance matrix (cross-assessment view)")
        print("  • Key Performance Indicators (KPIs)")
        print("  • Gap remediation tracker (consolidated)")
        print("  • Trend analysis over time")
        print("  • Evidence and approvals index")
        print("  • Pulls data from A.8.16.1 through A.8.16.4")
        print("\n  ⚠ This dashboard may show #REF errors until:")
        print("    1. All 4 source workbooks are completed")
        print("    2. Files are in same folder with correct names")
        print("    3. External links are updated in Excel")
        print("\n  Feynman's Law: The dashboard cannot hide what wasn't measured")
    
    else:
        print("\nGeneric Excel Workbook")
        print("  • No specific validation rules defined for this type")
        print("  • Standard Excel health checks applied")
    
    print("\n" + "=" * 80)


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("=" * 70)
        print("ISMS Control A.8.16 Monitoring Activities - Excel Sanity Checker")
        print("=" * 70)
        print("\nUsage: python3 excel_sanity_check_a816.py <filename.xlsx>")
        print("\nExamples:")
        print("  python3 excel_sanity_check_a816.py ISMS-IMP-A.8.16.1_Monitoring_Infrastructure_20260107.xlsx")
        print("  python3 excel_sanity_check_a816.py ISMS-IMP-A.8.16.2_Baseline_Detection_20260107.xlsx")
        print("  python3 excel_sanity_check_a816.py ISMS-IMP-A.8.16.3_Coverage_Assessment_20260107.xlsx")
        print("  python3 excel_sanity_check_a816.py ISMS-IMP-A.8.16.4_Alert_Management_20260107.xlsx")
        print("  python3 excel_sanity_check_a816.py ISMS-IMP-A.8.16.5_Compliance_Dashboard_20260107.xlsx")
        print("\nSupported workbook types:")
        print("  A.8.16.1 - Monitoring Infrastructure Assessment")
        print("  A.8.16.2 - Baseline & Detection Assessment")
        print("  A.8.16.3 - Coverage Assessment")
        print("  A.8.16.4 - Alert Management & Response Assessment")
        print("  A.8.16.5 - Compliance Dashboard")
        print("\n'You must not fool yourself' - Systems Engineering ISMS")
        sys.exit(1)
    
    filename = sys.argv[1]
    
    try:
        check_workbook_health(filename)
    except FileNotFoundError:
        print(f"\n❌ ERROR: File not found: {filename}")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION (syntax validated, structure verified)
# QA_TOOL: Claude Code Standardization
# =============================================================================
