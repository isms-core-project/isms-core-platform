#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.8.16.1 - Monitoring Infrastructure Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.16: Monitoring Activities
Assessment Domain 1 of 5: Monitoring Infrastructure and Platform Capabilities

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
monitoring infrastructure capabilities against ISO 27001:2022 Control A.8.16 
requirements.

**Purpose:**
Enables systematic assessment of monitoring platform capabilities, log source
coverage, data collection architecture, integration capabilities, and performance/
scalability against Control A.8.16 requirements.

**Assessment Scope:**
- Monitoring platform capabilities (SIEM/Log Management)
- Log source coverage and collection
- Data collection architecture and reliability
- Integration and enrichment capabilities
- Performance and scalability assessment
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and monitoring standards
2. Monitoring Platform - SIEM/platform capability assessment
3. Log Source Coverage - Comprehensive log source inventory and coverage
4. Data Collection Architecture - Collection methods and reliability
5. Integration & Enrichment - Context enrichment and correlation
6. Performance & Scale - Capacity and performance assessment
7. Summary Dashboard - Compliance metrics and gap summary
8. Evidence Register - Audit evidence tracking and documentation
9. Approval Sign-Off - Multi-level approval workflow

**Key Features:**
- Data validation with monitoring capability dropdown lists
- Conditional formatting for compliance status visualization
- Automated gap identification and remediation tracking
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with A.8.16 Compliance Dashboard

**Integration:**
This assessment feeds into ISMS-IMP-A.8.16.5 Compliance Dashboard, which
consolidates data from all five monitoring assessment domains for executive
oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a816_1_monitoring_infrastructure.py

Output:
    File: ISMS_A_8_16_1_Monitoring_Infrastructure_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Review monitoring platform capabilities documented
    2. Complete log source inventory (all systems generating logs)
    3. Assess data collection architecture and reliability
    4. Evaluate integration and enrichment capabilities
    5. Review performance and scalability metrics
    6. Conduct gap analysis for uncovered/unmonitored systems
    7. Define remediation actions with timelines
    8. Collect and link audit evidence
    9. Obtain stakeholder approvals
    10. Feed results into A.8.16.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.16
Assessment Domain:    1 of 5 (Monitoring Infrastructure)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization ISMS Team]
Date:                 24.01.2025
Last Modified:        24.01.2025
Python Version:       3.8+
License:              [Organization License/Terms]

Related Documents:
    - ISMS-POL-A.8.16: Monitoring Activities Policy (Governance)
    - ISMS-POL-A.8.16-S2.1: Monitoring Infrastructure Requirements
    - ISMS-IMP-A.8.16.1: Monitoring Infrastructure Implementation Guide
    - ISMS-IMP-A.8.16.2: Baseline & Detection Assessment (Domain 2)
    - ISMS-IMP-A.8.16.3: Coverage Assessment (Domain 3)
    - ISMS-IMP-A.8.16.4: Alert Management Assessment (Domain 4)
    - ISMS-IMP-A.8.16.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 24.01.2025
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.16.1 specification
    - Supports comprehensive monitoring infrastructure evaluation
    - Integrated with A.8.16.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Monitoring Philosophy:**
"The first principle is that you must not fool yourself — and you are the 
easiest person to fool." - Richard Feynman

This is not cargo cult ISMS. We assess CAPABILITIES, not checkbox compliance.
If your monitoring platform cannot detect threats, no policy will save you.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of monitoring capabilities and coverage.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- Monitoring platform architecture and capabilities
- Log sources and network topology
- Coverage gaps and security blind spots

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check for new systems requiring monitoring
- Semi-annually: Update coverage gaps and remediation status
- Annually: Complete reassessment of all monitoring infrastructure
- Ad-hoc: When infrastructure changes or new threats emerge

**Quality Assurance:**
Have SOC engineers and security architects validate assessments before using
results for compliance reporting or remediation decisions.

================================================================================
"""

import sys

# Dependency check
try:
    import openpyxl
except ImportError:
    print("\u274C Error: openpyxl not installed")
    print("ℹ️  Install with: sudo apt install python3-openpyxl")
    print("   or: pip3 install openpyxl")
    sys.exit(1)

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •



# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches ISMS-IMP-A.8.16.1 specification
    sheets = [
        "Instructions & Legend",
        "1. Monitoring Platform",
        "2. Log Source Coverage",
        "3. Data Collection Arch",
        "4. Integration Enrichment",
        "5. Performance Scale",
        "Summary Dashboard",
        "Evidence Register",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "example_row": {
            "font": Font(name="Calibri", size=9, italic=True, color="808080"),
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "border": border_thin,
        "status_compliant": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_noncompliant": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_na": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        },
        "gap_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "gap_high": {
            "fill": PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),
        },
        "gap_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "gap_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None,
            italic=style_dict["font"].italic if hasattr(style_dict["font"], 'italic') else False
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & COLUMN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    These are added to worksheet once, then applied to multiple cells.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_planned': DataValidation(
            type="list",
            formula1='"Yes,No,Planned"',
            allow_blank=False
        ),
        'yes_no_partial_na': DataValidation(
            type="list",
            formula1='"Yes,No,Partial,N/A"',
            allow_blank=False
        ),
        'yes_no_limited': DataValidation(
            type="list",
            formula1='"Yes,No,Limited"',
            allow_blank=False
        ),
        'platform_type': DataValidation(
            type="list",
            formula1='"SIEM,IDS/IPS,EDR,NDR,UEBA,Log Management,Other"',
            allow_blank=False
        ),
        'deployment_model': DataValidation(
            type="list",
            formula1='"On-Premises,Cloud,Hybrid"',
            allow_blank=False
        ),
        'quality_rating': DataValidation(
            type="list",
            formula1='"Excellent,Good,Limited,Poor"',
            allow_blank=False
        ),
        'search_performance': DataValidation(
            type="list",
            formula1='"<10 sec,10-60 sec,>60 sec"',
            allow_blank=False
        ),
        'ha_dr_status': DataValidation(
            type="list",
            formula1='"Documented,Tested,None"',
            allow_blank=False
        ),
        'implementation_status': DataValidation(
            type="list",
            formula1='"\u2705 Deployed,\u26A0\uFE0F Partial,\u274C Not Deployed,Planned"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1='"\u2705 Compliant,\u26A0\uFE0F Partial,\u274C Non-Compliant,N/A"',
            allow_blank=False
        ),
        'priority': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,None"',
            allow_blank=False
        ),
        'system_type': DataValidation(
            type="list",
            formula1='"Server,Network Device,Security Appliance,Endpoint,Cloud Service,Database,Application"',
            allow_blank=False
        ),
        'criticality': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'log_collection_status': DataValidation(
            type="list",
            formula1='"\u2705 Collecting,\u26A0\uFE0F Partial,\u274C Not Collecting,🔄 Planned"',
            allow_blank=False
        ),
        'collection_method': DataValidation(
            type="list",
            formula1='"Agent,Syslog,API,NetFlow,SNMP,WMI,Other"',
            allow_blank=False
        ),
        'protocol': DataValidation(
            type="list",
            formula1='"Syslog TCP,Syslog UDP,Syslog TLS,HTTP/S,API,Agent,SNMP,NetFlow,Other"',
            allow_blank=False
        ),
        'reliability': DataValidation(
            type="list",
            formula1='"High,Medium,Low,Unknown"',
            allow_blank=False
        ),
        'integration_type': DataValidation(
            type="list",
            formula1='"Real-time,Batch,Near Real-time,Manual"',
            allow_blank=False
        ),
        'enrichment_type': DataValidation(
            type="list",
            formula1='"Threat Intel,GeoIP,Asset Context,User Context,Custom,None"',
            allow_blank=False
        ),
    }
    
    # Add all validations to worksheet
    for val in validations.values():
        ws.add_data_validation(val)
    
    return validations


# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet with comprehensive guidance."""
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "MONITORING INFRASTRUCTURE ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 - Control A.8.16: Monitoring Activities"
    apply_style(ws["A2"], styles["subheader"])
    ws.row_dimensions[2].height = 25

    # Document Information Block
    row = 4
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    doc_info = [
        ("Document ID:", "ISMS-IMP-A.8.16.1"),
        ("Assessment Area:", "Monitoring Infrastructure"),
        ("Related Policy:", "ISMS-POL-A.8.16-S2.1"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT - DD.MM.YYYY]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Semi-annual"),
        ("Next Review Date:", "[Auto-calculated: Assessment Date + 6 months]"),
    ]
    
    row += 1
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if "USER INPUT" in value:
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Auto-calculate next review date
        if label == "Next Review Date:":
            ws[f"B{row}"] = '=IF(B8<>"","=B8+180","")'  # B8 is Assessment Date
        
        row += 1

    # How to Use This Workbook
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    instructions = [
        "1. Complete each worksheet tab in sequence (1-5)",
        "2. Use dropdown menus where provided - do not type directly",
        "3. Fill all yellow-highlighted cells with your organization's information",
        "4. Refer to reference tables on each sheet for guidance",
        "5. Complete compliance checklists on each assessment sheet",
        "6. Document exceptions in the Exception/Deviation blocks",
        "7. Gather evidence and list in Evidence Register",
        "8. Review Summary Dashboard for compliance gaps",
        "9. Obtain approvals via Approval Sign-Off sheet",
    ]

    row += 1
    for instruction in instructions:
        ws[f"A{row}"] = instruction
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 25
        row += 1

    # Legend Table
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "STATUS LEGEND"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    legend_headers = ["Symbol", "Meaning", "When to Use"]
    for col_idx, header in enumerate(legend_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    legend_data = [
        ("\u2705 Compliant", "Requirement fully met", "All controls in place"),
        ("\u26A0\uFE0F Partial", "Requirement partially met", "Some controls missing"),
        ("\u274C Non-Compliant", "Requirement not met", "No controls in place"),
        ("N/A", "Not applicable", "Requirement doesn't apply"),
    ]

    row += 1
    for symbol, meaning, when_to_use in legend_data:
        ws[f"A{row}"] = symbol
        ws[f"B{row}"] = meaning
        ws[f"C{row}"] = when_to_use
        
        # Apply conditional color coding
        if "Compliant" in symbol:
            ws[f"A{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif "Partial" in symbol:
            ws[f"A{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif "Non-Compliant" in symbol:
            ws[f"A{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif "N/A" in symbol:
            ws[f"A{row}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row += 1

    # Key Definitions
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "KEY DEFINITIONS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    definitions = [
        ("SIEM", "Security Information and Event Management - centralized log aggregation, correlation, and analysis"),
        ("IDS/IPS", "Intrusion Detection/Prevention System - network traffic monitoring for threats"),
        ("EDR", "Endpoint Detection and Response - endpoint-level threat monitoring"),
        ("NDR", "Network Detection and Response - network traffic analysis for threats"),
        ("UEBA", "User and Entity Behavior Analytics - anomaly detection based on behavior"),
        ("EPS", "Events Per Second - log ingestion rate capacity"),
        ("MTTA", "Mean Time to Acknowledge - average time to acknowledge an alert"),
        ("MTTD", "Mean Time to Detect - average time to detect a security incident"),
        ("MTTT", "Mean Time to Triage - average time to triage and classify an incident"),
        ("MTTR", "Mean Time to Resolve - average time to resolve a security incident"),
    ]

    row += 1
    ws[f"A{row}"] = "Term"
    ws[f"B{row}"] = "Definition"
    apply_style(ws[f"A{row}"], styles["column_header"])
    apply_style(ws[f"B{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:G{row}")

    row += 1
    for term, definition in definitions:
        ws[f"A{row}"] = term
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = definition
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:G{row}")
        ws.row_dimensions[row].height = 30
        row += 1

    # Feynman quote
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = '"The first principle is that you must not fool yourself — and you are the easiest person to fool." - Richard Feynman'
    ws[f"A{row}"].font = Font(italic=True, size=10, color="666666")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 35

    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "🎯 This is not cargo cult ISMS. This is evidence-based monitoring assessment."
    ws[f"A{row}"].font = Font(bold=True, size=10, color="003366")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    for col in ['D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15


# ============================================================================
# SECTION 4: SHEET 2 - MONITORING PLATFORM CAPABILITIES
# ============================================================================

def create_monitoring_platform_sheet(ws, styles):
    """Create Monitoring Platform Capabilities assessment sheet."""
    
    # Header
    ws.merge_cells("A1:U1")
    ws["A1"] = "1. MONITORING PLATFORM CAPABILITIES ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:U2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.4 - Assess SIEM/monitoring platform capabilities"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Does your organization have monitoring platforms (SIEM, IDS/IPS, EDR, NDR) with adequate capabilities?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:U3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 30

    # Response dropdown
    ws["A4"] = "Overall Status:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws.merge_cells("B4:E4")
    
    # Column Headers (Row 6)
    headers = [
        ("A", "Platform/Tool Name", 28),
        ("B", "Platform Type", 20),
        ("C", "Vendor/Solution", 22),
        ("D", "Deployment Model", 18),
        ("E", "Log Collection Methods", 24),
        ("F", "Parsing Capabilities", 20),
        ("G", "Storage & Indexing", 20),
        ("H", "Search Performance", 18),
        ("I", "Real-Time Alerting", 18),
        ("J", "Correlation Engine", 18),
        ("K", "Threat Intel Integration", 22),
        ("L", "SOAR Integration", 18),
        ("M", "Visualization/Dashboards", 20),
        ("N", "High Availability", 18),
        ("O", "Disaster Recovery", 18),
        ("P", "Current EPS Capacity", 18),
        ("Q", "Implementation Status", 20),
        ("R", "Last Upgrade Date", 16),
        ("S", "Compliance Status", 18),
        ("T", "Gaps/Issues", 30),
        ("U", "Remediation Priority", 18),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (Row 7)
    example_data = [
        "Splunk Enterprise",
        "SIEM",
        "Splunk Inc.",
        "Cloud",
        "Agents, Syslog, API",
        "Excellent",
        "Hot/Warm/Cold Tiers",
        "<10 sec",
        "Yes",
        "Advanced",
        "Yes",
        "Yes",
        "Excellent",
        "Yes",
        "Documented",
        "50,000 EPS",
        "\u2705 Deployed",
        "15.03.2024",
        "\u2705 Compliant",
        "None",
        "None"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-20: 13 rows)
    validations = create_base_validations(ws)
    
    for data_row in range(8, 21):
        for col_idx in range(1, 22):  # A to U
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            # Apply dropdowns based on column
            col_letter = get_column_letter(col_idx)
            if col_letter == 'B':  # Platform Type
                validations['platform_type'].add(cell)
            elif col_letter == 'D':  # Deployment Model
                validations['deployment_model'].add(cell)
            elif col_letter in ['F', 'M']:  # Quality ratings
                validations['quality_rating'].add(cell)
            elif col_letter == 'H':  # Search Performance
                validations['search_performance'].add(cell)
            elif col_letter in ['I', 'J', 'K', 'L', 'N']:  # Yes/No/Planned
                validations['yes_no_planned'].add(cell)
            elif col_letter == 'O':  # DR Status
                validations['ha_dr_status'].add(cell)
            elif col_letter == 'Q':  # Implementation Status
                validations['implementation_status'].add(cell)
            elif col_letter == 'S':  # Compliance Status
                validations['compliance_status'].add(cell)
            elif col_letter == 'U':  # Priority
                validations['priority'].add(cell)

    # Compliance Checklist (Starting Row 22)
    row = 22
    ws.merge_cells(f"A{row}:U{row}")
    ws[f"A{row}"] = "MONITORING PLATFORM CAPABILITIES CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Primary monitoring platform (SIEM) deployed",
        "Log collection supports multiple methods (agents, syslog, API)",
        "Parsing capabilities for common log formats (Windows, Linux, network devices)",
        "Storage includes hot/warm/cold tiers for cost optimization",
        "Search performance meets requirements (<10 sec for typical queries)",
        "Real-time alerting capability enabled",
        "Correlation engine supports multi-event detection",
        "Threat intelligence integration implemented",
        "SOAR or ticketing integration functional",
        "Dashboards configured for SOC operations",
        "High availability architecture implemented",
        "Disaster recovery plan documented and tested",
        "Platform capacity adequate for current log volume",
        "Platform upgraded within last 12 months",
        "Platform security hardened (RBAC, MFA, encryption)",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.merge_cells(f"B{row}:T{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:T{row}")
        
        # Status dropdown
        ws[f"U{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['compliance_status'].add(ws[f"U{row}"])
        
        ws.row_dimensions[row].height = 25
        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(U23:U37,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Reference Tables (Starting Row 40)
    row = 40
    ws.merge_cells(f"A{row}:U{row}")
    ws[f"A{row}"] = "REFERENCE TABLES"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    # Table 1: Monitoring Platform Types
    row += 2
    ws[f"A{row}"] = "Table 1: Monitoring Platform Types"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="003366")
    ws.merge_cells(f"A{row}:U{row}")

    row += 1
    table1_headers = ["Platform Type", "Primary Purpose", "Example Capabilities"]
    for col_idx, header in enumerate(table1_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    table1_data = [
        ("SIEM", "Centralized log aggregation & correlation", "Search, alert, correlate, dashboards"),
        ("IDS/IPS", "Network intrusion detection", "Signature matching, anomaly detection"),
        ("EDR", "Endpoint threat detection", "Process monitoring, behavioral analysis"),
        ("NDR", "Network traffic analysis", "Flow analysis, encrypted traffic analysis"),
        ("UEBA", "User behavior analytics", "Baseline deviation, risk scoring"),
    ]

    row += 1
    for platform, purpose, capabilities in table1_data:
        ws[f"A{row}"] = platform
        ws[f"B{row}"] = purpose
        ws[f"C{row}"] = capabilities
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:N{row}")
        ws.merge_cells(f"C{row}:U{row}")
        ws.row_dimensions[row].height = 30
        row += 1

    # Table 2: Critical SIEM Capabilities
    row += 2
    ws[f"A{row}"] = "Table 2: Critical SIEM Capabilities (Per S2.1.4.1)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="003366")
    ws.merge_cells(f"A{row}:U{row}")

    row += 1
    table2_headers = ["Capability", "Requirement Level", "Validation Method"]
    for col_idx, header in enumerate(table2_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    table2_data = [
        ("Multi-format log parsing", "MUST HAVE", "Test with sample logs"),
        ("Indexed storage", "MUST HAVE", "Verify search performance"),
        ("Real-time alerting", "MUST HAVE", "Test alert generation"),
        ("Multi-event correlation", "MUST HAVE", "Create test correlation rule"),
        ("Threat intel integration", "MUST HAVE", "Verify IOC matching"),
        ("Search < 10 sec", "SHOULD HAVE", "Benchmark queries"),
    ]

    row += 1
    for capability, level, validation in table2_data:
        ws[f"A{row}"] = capability
        ws[f"B{row}"] = level
        ws[f"C{row}"] = validation
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:H{row}")
        ws.merge_cells(f"B{row}:N{row}")
        ws.merge_cells(f"C{row}:U{row}")
        ws.row_dimensions[row].height = 25
        row += 1

    # Exception/Deviation Block (Starting Row 70)
    row = 70
    ws.merge_cells(f"A{row}:U{row}")
    ws[f"A{row}"] = "EXCEPTION/DEVIATION MANAGEMENT"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    exception_headers = [
        "Exception ID", "Risk ID", "Business Justification", 
        "Compensating Controls", "Approval Status", "Review Date"
    ]
    for col_idx, header in enumerate(exception_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    # 5 exception rows
    for i in range(5):
        row += 1
        for col_idx in range(1, 7):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.row_dimensions[row].height = 25

# ============================================================================
# SECTION 5: SHEET 3 - LOG SOURCE COVERAGE ASSESSMENT
# ============================================================================

def create_log_source_coverage_sheet(ws, styles):
    """Create Log Source Coverage assessment sheet."""
    
    # Header
    ws.merge_cells("A1:T1")
    ws["A1"] = "2. LOG SOURCE COVERAGE ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:T2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.2 - Assess log source coverage"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Does your organization collect logs from all critical systems and adequate coverage of standard systems?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:T3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 30

    # Response dropdown
    ws["A4"] = "Overall Coverage Status:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws.merge_cells("B4:E4")

    # Column Headers (Row 6)
    headers = [
        ("A", "System/Asset Name", 28),
        ("B", "System Type", 22),
        ("C", "Criticality", 15),
        ("D", "Location", 18),
        ("E", "System Owner", 20),
        ("F", "OS/Platform", 18),
        ("G", "Log Types Collected", 30),
        ("H", "Collection Method", 20),
        ("I", "Collection Protocol", 20),
        ("J", "Log Volume (GB/day)", 18),
        ("K", "Retention Period (days)", 20),
        ("L", "Parsing Status", 18),
        ("M", "Integration with SIEM", 20),
        ("N", "Collection Status", 20),
        ("O", "Last Verified", 16),
        ("P", "Reliability", 15),
        ("Q", "Compliance Status", 18),
        ("R", "Gaps/Issues", 30),
        ("S", "Remediation Plan", 30),
        ("T", "Priority", 15),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (Row 7)
    example_data = [
        "DC01-WinServer",
        "Server",
        "Critical",
        "Primary DC",
        "IT Operations",
        "Windows Server 2022",
        "Security, System, Application",
        "Agent",
        "HTTPS",
        "2.5",
        "365",
        "Parsed",
        "Yes",
        "\u2705 Collecting",
        "15.12.2024",
        "High",
        "\u2705 Compliant",
        "None",
        "N/A",
        "None"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-40: 33 rows for comprehensive coverage)
    validations = create_base_validations(ws)
    
    for data_row in range(8, 41):
        for col_idx in range(1, 21):  # A to T
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            # Apply dropdowns based on column
            col_letter = get_column_letter(col_idx)
            if col_letter == 'B':  # System Type
                validations['system_type'].add(cell)
            elif col_letter == 'C':  # Criticality
                validations['criticality'].add(cell)
            elif col_letter == 'H':  # Collection Method
                validations['collection_method'].add(cell)
            elif col_letter == 'I':  # Protocol
                validations['protocol'].add(cell)
            elif col_letter == 'N':  # Collection Status
                validations['log_collection_status'].add(cell)
            elif col_letter == 'P':  # Reliability
                validations['reliability'].add(cell)
            elif col_letter == 'Q':  # Compliance Status
                validations['compliance_status'].add(cell)
            elif col_letter == 'T':  # Priority
                validations['priority'].add(cell)

    # Coverage Summary Statistics (Starting Row 43)
    row = 43
    ws.merge_cells(f"A{row}:T{row}")
    ws[f"A{row}"] = "LOG SOURCE COVERAGE STATISTICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"B{row}"] = "Count/Value"
    ws[f"C{row}"] = "Target"
    ws[f"D{row}"] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:G{row}")
    ws.merge_cells(f"B{row}:M{row}")
    ws.merge_cells(f"C{row}:P{row}")
    ws.merge_cells(f"D{row}:T{row}")

    statistics = [
        ("Total Systems Inventoried", "=COUNTA(A8:A40)", "N/A", ""),
        ("Critical Systems with Logs", "=COUNTIFS(C8:C40,\"Critical\",N8:N40,\"\u2705 Collecting\")", "100%", ""),
        ("High Priority Systems with Logs", "=COUNTIFS(C8:C40,\"High\",N8:N40,\"\u2705 Collecting\")", ">90%", ""),
        ("Total Log Volume (GB/day)", "=SUM(J8:J40)", "Monitor", ""),
        ("Systems with Compliance Gaps", "=COUNTIFS(Q8:Q40,\"\u274C Non-Compliant\")+COUNTIFS(Q8:Q40,\"\u26A0\uFE0F Partial\")", "0", ""),
        ("Collection Reliability High", "=COUNTIF(P8:P40,\"High\")", "Target", ""),
        ("Collection Reliability Low", "=COUNTIF(P8:P40,\"Low\")", "0", ""),
    ]

    row += 1
    for metric, formula, target, status in statistics:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = formula
        ws[f"C{row}"] = target
        ws[f"D{row}"] = status
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:G{row}")
        ws.merge_cells(f"B{row}:M{row}")
        ws.merge_cells(f"C{row}:P{row}")
        ws.merge_cells(f"D{row}:T{row}")
        ws.row_dimensions[row].height = 25
        row += 1

    # Compliance Checklist (Starting Row 53)
    row = 53
    ws.merge_cells(f"A{row}:T{row}")
    ws[f"A{row}"] = "LOG SOURCE COVERAGE CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "All critical systems (servers, databases) generating logs",
        "All network devices (firewalls, routers, switches) generating logs",
        "All security appliances (IDS/IPS, proxies) generating logs",
        "Authentication systems (AD, LDAP, SSO) generating logs",
        "Cloud services generating logs (AWS, Azure, O365)",
        "Critical applications generating logs",
        "Endpoints (workstations, laptops) generating logs where required",
        "Log collection methods reliable (agents, syslog, API)",
        "Log volume monitored and within capacity",
        "Log retention periods meet policy requirements (min 90 days for critical)",
        "Parsing configured for all critical log sources",
        "SIEM integration functional for all critical sources",
        "Log collection verified within last 30 days",
        "Gaps documented with remediation plans",
        "Coverage assessed quarterly",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:S{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:S{row}")
        
        # Status dropdown
        ws[f"T{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['compliance_status'].add(ws[f"T{row}"])
        
        ws.row_dimensions[row].height = 25
        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(T55:T69,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Reference Table: Critical System Types
    row = 75
    ws.merge_cells(f"A{row}:T{row}")
    ws[f"A{row}"] = "REFERENCE: CRITICAL SYSTEM TYPES REQUIRING LOG COLLECTION"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="003366")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ref_headers = ["System Type", "Required Log Types", "Minimum Retention"]
    for col_idx, header in enumerate(ref_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    ref_data = [
        ("Domain Controllers", "Security, System, Directory Service", "365 days"),
        ("Database Servers", "Audit, Error, Query logs", "365 days"),
        ("Web Servers", "Access, Error, Security", "180 days"),
        ("Firewalls", "Traffic, Deny, Alert", "365 days"),
        ("VPN Gateways", "Authentication, Connection, Error", "180 days"),
        ("Email Servers", "Message tracking, Security", "180 days"),
        ("Cloud Services (AWS/Azure/O365)", "Activity, Security, Configuration", "365 days"),
    ]

    row += 1
    for system_type, log_types, retention in ref_data:
        ws[f"A{row}"] = system_type
        ws[f"B{row}"] = log_types
        ws[f"C{row}"] = retention
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:F{row}")
        ws.merge_cells(f"B{row}:N{row}")
        ws.merge_cells(f"C{row}:T{row}")
        ws.row_dimensions[row].height = 30
        row += 1


# ============================================================================
# SECTION 6: SHEET 4 - DATA COLLECTION ARCHITECTURE
# ============================================================================

def create_data_collection_architecture_sheet(ws, styles):
    """Create Data Collection Architecture assessment sheet."""
    
    # Header
    ws.merge_cells("A1:R1")
    ws["A1"] = "3. DATA COLLECTION ARCHITECTURE ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:R2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.3 - Assess data collection architecture"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Is your log collection architecture reliable, secure, and scalable?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:R3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 30

    # Column Headers (Row 6)
    headers = [
        ("A", "Collection Component", 28),
        ("B", "Component Type", 22),
        ("C", "Purpose/Function", 30),
        ("D", "Protocol Used", 20),
        ("E", "Encryption Status", 18),
        ("F", "Authentication Method", 22),
        ("G", "Throughput Capacity", 18),
        ("H", "Current Utilization %", 18),
        ("I", "Buffer/Queue Size", 18),
        ("J", "Failover Configured", 18),
        ("K", "Load Balancing", 18),
        ("L", "Monitoring Enabled", 18),
        ("M", "Health Check Interval", 18),
        ("N", "Last Maintenance", 16),
        ("O", "Implementation Status", 20),
        ("P", "Compliance Status", 18),
        ("Q", "Issues/Gaps", 30),
        ("R", "Priority", 15),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (Row 7)
    example_data = [
        "Syslog Forwarder Cluster",
        "Log Forwarder",
        "Central syslog collection",
        "Syslog TLS",
        "TLS 1.3",
        "Certificate-based",
        "100K EPS",
        "65%",
        "10GB",
        "Yes",
        "Yes",
        "Yes",
        "60 seconds",
        "10.12.2024",
        "\u2705 Deployed",
        "\u2705 Compliant",
        "None",
        "None"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-25: 18 rows)
    validations = create_base_validations(ws)
    
    for data_row in range(8, 26):
        for col_idx in range(1, 19):  # A to R
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            # Apply dropdowns based on column
            col_letter = get_column_letter(col_idx)
            if col_letter == 'D':  # Protocol
                validations['protocol'].add(cell)
            elif col_letter in ['J', 'K', 'L']:  # Yes/No
                validations['yes_no'].add(cell)
            elif col_letter == 'O':  # Implementation Status
                validations['implementation_status'].add(cell)
            elif col_letter == 'P':  # Compliance Status
                validations['compliance_status'].add(cell)
            elif col_letter == 'R':  # Priority
                validations['priority'].add(cell)

    # Architecture Compliance Checklist (Starting Row 28)
    row = 28
    ws.merge_cells(f"A{row}:R{row}")
    ws[f"A{row}"] = "DATA COLLECTION ARCHITECTURE CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Log collection uses encrypted protocols (TLS/HTTPS)",
        "Authentication required for log submission",
        "Collection components have adequate throughput capacity",
        "Buffering/queuing configured to prevent log loss",
        "Failover mechanisms implemented for critical components",
        "Load balancing configured where applicable",
        "Collection infrastructure monitored (health, performance)",
        "Alerts configured for collection failures",
        "Collection reliability >99% for critical systems",
        "Time synchronization (NTP) configured across infrastructure",
        "Log integrity mechanisms in place (checksums, signatures)",
        "Collection infrastructure documented (architecture diagrams)",
        "Capacity planning performed quarterly",
        "Disaster recovery tested for collection infrastructure",
        "Access to collection infrastructure restricted (RBAC)",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:Q{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:Q{row}")
        
        # Status dropdown
        ws[f"R{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['compliance_status'].add(ws[f"R{row}"])
        
        ws.row_dimensions[row].height = 25
        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(R30:R44,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Reference Table: Collection Protocols Security
    row = 50
    ws.merge_cells(f"A{row}:R{row}")
    ws[f"A{row}"] = "REFERENCE: COLLECTION PROTOCOL SECURITY LEVELS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="003366")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ref_headers = ["Protocol", "Security Level", "Use Case", "Recommendations"]
    for col_idx, header in enumerate(ref_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    ref_data = [
        ("Syslog TLS (TCP 6514)", "High", "Secure syslog", "Recommended for all external collection"),
        ("HTTPS/API", "High", "Cloud services, REST APIs", "Use for cloud and API integrations"),
        ("Agent-based (HTTPS)", "High", "Endpoint/server collection", "Preferred for servers and endpoints"),
        ("Syslog TCP (TCP 514)", "Medium", "Legacy devices", "Use only when TLS unavailable"),
        ("Syslog UDP (UDP 514)", "Low", "Legacy, low-priority", "Avoid for critical systems"),
        ("SNMP v3", "Medium", "Network device monitoring", "Acceptable for network devices"),
        ("NetFlow", "Medium", "Network traffic metadata", "Supplementary to syslog"),
    ]

    row += 1
    for protocol, security, use_case, recommendation in ref_data:
        ws[f"A{row}"] = protocol
        ws[f"B{row}"] = security
        ws[f"C{row}"] = use_case
        ws[f"D{row}"] = recommendation
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"D{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:E{row}")
        ws.merge_cells(f"B{row}:G{row}")
        ws.merge_cells(f"C{row}:J{row}")
        ws.merge_cells(f"D{row}:R{row}")
        ws.row_dimensions[row].height = 30
        row += 1


# ============================================================================
# SECTION 7: SHEET 5 - INTEGRATION & ENRICHMENT
# ============================================================================

def create_integration_enrichment_sheet(ws, styles):
    """Create Integration & Enrichment assessment sheet."""
    
    # Header
    ws.merge_cells("A1:P1")
    ws["A1"] = "4. INTEGRATION & ENRICHMENT ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:P2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.5 - Assess integration and data enrichment"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Are monitoring systems properly integrated with threat intelligence, asset inventory, and other context sources?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:P3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 30

    # Column Headers (Row 6)
    headers = [
        ("A", "Integration/Enrichment Name", 30),
        ("B", "Integration Type", 22),
        ("C", "Data Source", 25),
        ("D", "Enrichment Type", 22),
        ("E", "Integration Method", 20),
        ("F", "Update Frequency", 18),
        ("G", "Data Quality", 18),
        ("H", "Coverage %", 15),
        ("I", "Latency", 15),
        ("J", "Reliability", 15),
        ("K", "Last Updated", 16),
        ("L", "Implementation Status", 20),
        ("M", "Compliance Status", 18),
        ("N", "Value/Impact", 25),
        ("O", "Issues/Gaps", 30),
        ("P", "Priority", 15),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (Row 7)
    example_data = [
        "MISP Threat Feed",
        "Real-time",
        "MISP Community",
        "Threat Intel",
        "API",
        "Hourly",
        "Excellent",
        "95%",
        "<1 min",
        "High",
        "07.01.2025",
        "\u2705 Deployed",
        "\u2705 Compliant",
        "IOC matching for alerts",
        "None",
        "High"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-22: 15 rows)
    validations = create_base_validations(ws)
    
    for data_row in range(8, 23):
        for col_idx in range(1, 17):  # A to P
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            # Apply dropdowns based on column
            col_letter = get_column_letter(col_idx)
            if col_letter == 'B':  # Integration Type
                validations['integration_type'].add(cell)
            elif col_letter == 'D':  # Enrichment Type
                validations['enrichment_type'].add(cell)
            elif col_letter == 'G':  # Data Quality
                validations['quality_rating'].add(cell)
            elif col_letter == 'J':  # Reliability
                validations['reliability'].add(cell)
            elif col_letter == 'L':  # Implementation Status
                validations['implementation_status'].add(cell)
            elif col_letter == 'M':  # Compliance Status
                validations['compliance_status'].add(cell)
            elif col_letter == 'P':  # Priority
                validations['priority'].add(cell)

    # Integration Compliance Checklist (Starting Row 25)
    row = 25
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "INTEGRATION & ENRICHMENT CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "Threat intelligence feed(s) integrated into SIEM",
        "Threat intelligence updated at least daily",
        "GeoIP enrichment configured for IP addresses",
        "Asset inventory integrated (hostname, owner, criticality)",
        "User context enrichment (AD/LDAP integration)",
        "DNS enrichment for domain resolution",
        "CMDB integration for asset context",
        "Vulnerability data integrated (scanner results)",
        "Ticketing system integration (ServiceNow, JIRA)",
        "SOAR platform integration for automated response",
        "Cloud provider APIs integrated (AWS, Azure, GCP)",
        "Enrichment latency acceptable (<5 minutes)",
        "Enrichment coverage >90% for critical assets",
        "Integration health monitored and alerted",
        "Integration documentation maintained",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:O{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:O{row}")
        
        # Status dropdown
        ws[f"P{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['compliance_status'].add(ws[f"P{row}"])
        
        ws.row_dimensions[row].height = 25
        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(P27:P41,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

    # Reference Table: Common Enrichment Sources
    row = 47
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "REFERENCE: COMMON ENRICHMENT SOURCES"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="003366")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ref_headers = ["Enrichment Type", "Example Sources", "Benefit to Monitoring"]
    for col_idx, header in enumerate(ref_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])

    ref_data = [
        ("Threat Intelligence", "MISP, AlienVault OTX, VirusTotal, Abuse.ch", "IOC detection, malicious IP/domain identification"),
        ("GeoIP", "MaxMind, IP2Location", "Geo-based anomaly detection, compliance"),
        ("Asset Context", "CMDB, ServiceNow, Lansweeper", "Criticality assessment, owner identification"),
        ("User Context", "Active Directory, Okta, Azure AD", "User behavior analysis, privilege context"),
        ("Vulnerability Data", "Qualys, Tenable, Rapid7", "Risk scoring, exploit correlation"),
        ("DNS", "Internal DNS, passive DNS", "Domain reputation, DGA detection"),
    ]

    row += 1
    for enrich_type, sources, benefit in ref_data:
        ws[f"A{row}"] = enrich_type
        ws[f"B{row}"] = sources
        ws[f"C{row}"] = benefit
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:D{row}")
        ws.merge_cells(f"B{row}:H{row}")
        ws.merge_cells(f"C{row}:P{row}")
        ws.row_dimensions[row].height = 30
        row += 1


# ============================================================================
# SECTION 8: SHEET 6 - PERFORMANCE & SCALABILITY
# ============================================================================

def create_performance_scalability_sheet(ws, styles):
    """Create Performance & Scalability assessment sheet."""
    
    # Header
    ws.merge_cells("A1:Q1")
    ws["A1"] = "5. PERFORMANCE & SCALABILITY ASSESSMENT"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:Q2")
    ws["A2"] = "Policy Reference: ISMS-POL-A.8.16-S2.1.6 - Assess monitoring performance and scalability"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment Question
    ws["A3"] = "Is the monitoring infrastructure performing adequately and scalable for future growth?"
    ws["A3"].font = Font(bold=True, size=11)
    ws.merge_cells("A3:Q3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 30

    # Performance Metrics Section (Starting Row 5)
    row = 5
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "CURRENT PERFORMANCE METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    # Column Headers (Row 6)
    headers = [
        ("A", "Metric", 35),
        ("B", "Current Value", 20),
        ("C", "Unit", 15),
        ("D", "Target/Threshold", 20),
        ("E", "Status", 18),
        ("F", "Trend", 18),
        ("G", "Last Measured", 18),
        ("H", "Notes", 35),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Performance metrics data rows
    performance_metrics = [
        ("SIEM Ingestion Rate (Current)", "", "EPS", "N/A", "", "", "", ""),
        ("SIEM Ingestion Rate (Peak)", "", "EPS", "N/A", "", "", "", ""),
        ("SIEM Licensed Capacity", "", "EPS", "N/A", "", "", "", ""),
        ("Capacity Utilization", "", "%", "<80%", "", "", "", ""),
        ("Average Search Response Time", "", "seconds", "<10 sec", "", "", "", ""),
        ("Indexing Lag", "", "minutes", "<5 min", "", "", "", ""),
        ("Storage Utilization", "", "%", "<80%", "", "", "", ""),
        ("Hot Tier Storage Available", "", "TB", ">20%", "", "", "", ""),
        ("Daily Log Volume", "", "GB/day", "N/A", "", "", "", ""),
        ("Data Retention Compliance", "", "days", ">=90", "", "", "", ""),
        ("SIEM Availability (Last 30d)", "", "%", ">99.5%", "", "", "", ""),
        ("Alert Generation Latency", "", "seconds", "<60 sec", "", "", "", ""),
        ("Correlation Rule Execution Time", "", "seconds", "<5 sec", "", "", "", ""),
        ("Dashboard Load Time", "", "seconds", "<3 sec", "", "", "", ""),
        ("API Response Time", "", "milliseconds", "<500 ms", "", "", "", ""),
    ]

    validations = create_base_validations(ws)
    row = 7
    for metric, value, unit, target, status, trend, measured, notes in performance_metrics:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = value
        ws[f"C{row}"] = unit
        ws[f"D{row}"] = target
        ws[f"E{row}"] = status
        ws[f"F{row}"] = trend
        ws[f"G{row}"] = measured
        ws[f"H{row}"] = notes
        
        # Style first column
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        
        # Yellow input cells for current value and notes
        for col in ['B', 'F', 'G', 'H']:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Status dropdown
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['compliance_status'].add(ws[f"E{row}"])
        
        ws.merge_cells(f"H{row}:Q{row}")
        ws.row_dimensions[row].height = 25
        row += 1

    # Scalability Assessment (Starting Row 24)
    row = 24
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "SCALABILITY ASSESSMENT"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    scalability_questions = [
        ("Current log sources (count)", "", "N/A"),
        ("Projected log sources (12 months)", "", "N/A"),
        ("Current daily log volume (GB)", "", "N/A"),
        ("Projected daily log volume (12 months, GB)", "", "N/A"),
        ("Growth rate (%)", "", "<30% increase"),
        ("Available capacity for growth (%)", "", ">50%"),
        ("Time to add new log source", "", "<1 hour"),
        ("Horizontal scaling capability", "", "Yes"),
        ("Vertical scaling capability", "", "Yes"),
        ("Capacity planning performed", "", "Quarterly"),
    ]

    row += 1
    ws[f"A{row}"] = "Scalability Factor"
    ws[f"B{row}"] = "Current/Projected"
    ws[f"C{row}"] = "Target"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"A{row}:H{row}")
    ws.merge_cells(f"B{row}:M{row}")
    ws.merge_cells(f"C{row}:Q{row}")

    row += 1
    for factor, value, target in scalability_questions:
        ws[f"A{row}"] = factor
        ws[f"B{row}"] = value
        ws[f"C{row}"] = target
        
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        ws.merge_cells(f"A{row}:H{row}")
        ws.merge_cells(f"B{row}:M{row}")
        ws.merge_cells(f"C{row}:Q{row}")
        ws.row_dimensions[row].height = 25
        row += 1

    # Performance Checklist (Starting Row 37)
    row = 37
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "PERFORMANCE & SCALABILITY CHECKLIST"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "SIEM capacity utilization <80%",
        "Search performance meets <10 second target",
        "Indexing lag <5 minutes",
        "Storage utilization <80%",
        "SIEM availability >99.5%",
        "Alert latency <60 seconds",
        "Capacity planning performed quarterly",
        "Growth projections documented",
        "Scalability tested (load testing)",
        "Performance baselines established",
        "Performance monitoring automated",
        "Performance metrics reviewed monthly",
        "Capacity alerts configured",
        "Retention policies optimized for cost",
        "Disaster recovery capacity validated",
    ]

    row += 1
    ws[f"A{row}"] = "#"
    ws[f"B{row}"] = "Requirement"
    ws[f"C{row}"] = "Status"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:P{row}")

    row += 1
    for idx, item in enumerate(checklist_items, start=1):
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:P{row}")
        
        # Status dropdown
        ws[f"Q{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['compliance_status'].add(ws[f"Q{row}"])
        
        ws.row_dimensions[row].height = 25
        row += 1

    # Auto-score formula
    score_row = row
    ws[f"A{score_row}"] = "SCORE:"
    ws[f"A{score_row}"].font = Font(bold=True, size=12)
    ws[f"B{score_row}"] = f'=COUNTIF(Q39:Q53,"\u2705 Compliant")&" / 15"'
    ws[f"B{score_row}"].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f"B{score_row}:E{score_row}")

# ============================================================================
# SECTION 9: SHEET 7 - SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create Summary Dashboard with overall compliance view."""
    
    # Header
    ws.merge_cells("A1:P1")
    ws["A1"] = "MONITORING INFRASTRUCTURE - SUMMARY DASHBOARD"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:P2")
    ws["A2"] = "Consolidated Assessment Overview - ISMS-IMP-A.8.16.1"
    apply_style(ws["A2"], styles["subheader"])
    ws.row_dimensions[2].height = 25

    # Overall Assessment Status (Row 4)
    row = 4
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "OVERALL ASSESSMENT STATUS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    # Summary Statistics (Rows 6-15)
    row = 6
    ws[f"A{row}"] = "Assessment Area"
    ws[f"B{row}"] = "Compliant"
    ws[f"C{row}"] = "Partial"
    ws[f"D{row}"] = "Non-Compliant"
    ws[f"E{row}"] = "N/A"
    ws[f"F{row}"] = "Total Items"
    ws[f"G{row}"] = "Compliance %"
    ws[f"H{row}"] = "Status"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])

    summary_areas = [
        ("1. Monitoring Platform Capabilities", 
         '=COUNTIF(\'1. Monitoring Platform\'!U23:U37,"\u2705 Compliant")',
         '=COUNTIF(\'1. Monitoring Platform\'!U23:U37,"\u26A0\uFE0F Partial")',
         '=COUNTIF(\'1. Monitoring Platform\'!U23:U37,"\u274C Non-Compliant")',
         '=COUNTIF(\'1. Monitoring Platform\'!U23:U37,"N/A")',
         "15"),
        ("2. Log Source Coverage",
         '=COUNTIF(\'2. Log Source Coverage\'!T55:T69,"\u2705 Compliant")',
         '=COUNTIF(\'2. Log Source Coverage\'!T55:T69,"\u26A0\uFE0F Partial")',
         '=COUNTIF(\'2. Log Source Coverage\'!T55:T69,"\u274C Non-Compliant")',
         '=COUNTIF(\'2. Log Source Coverage\'!T55:T69,"N/A")',
         "15"),
        ("3. Data Collection Architecture",
         '=COUNTIF(\'3. Data Collection Arch\'!R30:R44,"\u2705 Compliant")',
         '=COUNTIF(\'3. Data Collection Arch\'!R30:R44,"\u26A0\uFE0F Partial")',
         '=COUNTIF(\'3. Data Collection Arch\'!R30:R44,"\u274C Non-Compliant")',
         '=COUNTIF(\'3. Data Collection Arch\'!R30:R44,"N/A")',
         "15"),
        ("4. Integration & Enrichment",
         '=COUNTIF(\'4. Integration Enrichment\'!P27:P41,"\u2705 Compliant")',
         '=COUNTIF(\'4. Integration Enrichment\'!P27:P41,"\u26A0\uFE0F Partial")',
         '=COUNTIF(\'4. Integration Enrichment\'!P27:P41,"\u274C Non-Compliant")',
         '=COUNTIF(\'4. Integration Enrichment\'!P27:P41,"N/A")',
         "15"),
        ("5. Performance & Scalability",
         '=COUNTIF(\'5. Performance Scale\'!Q39:Q53,"\u2705 Compliant")',
         '=COUNTIF(\'5. Performance Scale\'!Q39:Q53,"\u26A0\uFE0F Partial")',
         '=COUNTIF(\'5. Performance Scale\'!Q39:Q53,"\u274C Non-Compliant")',
         '=COUNTIF(\'5. Performance Scale\'!Q39:Q53,"N/A")',
         "15"),
    ]

    row = 7
    for area, compliant, partial, non_compliant, na, total in summary_areas:
        ws[f"A{row}"] = area
        ws[f"B{row}"] = compliant
        ws[f"C{row}"] = partial
        ws[f"D{row}"] = non_compliant
        ws[f"E{row}"] = na
        ws[f"F{row}"] = total
        ws[f"G{row}"] = f'=IF(F{row}>0,ROUND(B{row}/(F{row}-E{row})*100,1)&"%","N/A")'
        ws[f"H{row}"] = f'=IF(G{row}="N/A","N/A",IF(VALUE(LEFT(G{row},LEN(G{row})-1))>=80,"\u2705 Good",IF(VALUE(LEFT(G{row},LEN(G{row})-1))>=60,"\u26A0\uFE0F Needs Work","\u274C Critical")))'
        
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A{row}:A{row}")
        ws.row_dimensions[row].height = 25
        row += 1

    # Overall Totals
    row += 1
    ws[f"A{row}"] = "OVERALL TOTAL"
    ws[f"B{row}"] = "=SUM(B7:B11)"
    ws[f"C{row}"] = "=SUM(C7:C11)"
    ws[f"D{row}"] = "=SUM(D7:D11)"
    ws[f"E{row}"] = "=SUM(E7:E11)"
    ws[f"F{row}"] = "=SUM(F7:F11)"
    ws[f"G{row}"] = f'=IF(F{row}>0,ROUND(B{row}/(F{row}-E{row})*100,1)&"%","N/A")'
    ws[f"H{row}"] = f'=IF(G{row}="N/A","N/A",IF(VALUE(LEFT(G{row},LEN(G{row})-1))>=80,"\u2705 Good",IF(VALUE(LEFT(G{row},LEN(G{row})-1))>=60,"\u26A0\uFE0F Needs Work","\u274C Critical")))'
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f"{col}{row}"].font = Font(bold=True, size=11)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Key Findings Section (Starting Row 16)
    row = 16
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "KEY FINDINGS & RECOMMENDATIONS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Category"
    ws[f"B{row}"] = "Finding/Recommendation"
    ws[f"C{row}"] = "Priority"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 15

    # 10 rows for key findings
    validations = create_base_validations(ws)
    for i in range(10):
        row += 1
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:N{row}")
        ws.merge_cells(f"C{row}:P{row}")
        validations['priority'].add(ws[f"C{row}"])
        ws.row_dimensions[row].height = 35

    # Critical Gaps Summary (Starting Row 30)
    row = 30
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "CRITICAL GAPS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Gap ID"
    ws[f"B{row}"] = "Description"
    ws[f"C{row}"] = "Impact"
    ws[f"D{row}"] = "Owner"
    ws[f"E{row}"] = "Target Date"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f"{col}{row}"], styles["column_header"])
    ws.merge_cells(f"B{row}:K{row}")
    ws.merge_cells(f"C{row}:M{row}")
    ws.merge_cells(f"D{row}:N{row}")
    ws.merge_cells(f"E{row}:P{row}")

    # 8 rows for critical gaps
    for i in range(8):
        row += 1
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"B{row}:K{row}")
        ws.merge_cells(f"C{row}:M{row}")
        ws.merge_cells(f"D{row}:N{row}")
        ws.merge_cells(f"E{row}:P{row}")
        ws.row_dimensions[row].height = 30

    # Assessor Notes (Starting Row 42)
    row = 42
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "ASSESSOR NOTES & OBSERVATIONS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws.merge_cells(f"A{row}:P{row+10}")
    ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    for i in range(11):
        ws.row_dimensions[row + i].height = 25


# ============================================================================
# SECTION 10: SHEET 8 - EVIDENCE REGISTER
# ============================================================================

def create_evidence_register_sheet(ws, styles):
    """Create Evidence Register for audit documentation."""
    
    # Header
    ws.merge_cells("A1:L1")
    ws["A1"] = "EVIDENCE REGISTER"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:L2")
    ws["A2"] = "Evidence Collection and Documentation - ISMS-IMP-A.8.16.1"
    apply_style(ws["A2"], styles["subheader"])
    ws.row_dimensions[2].height = 25

    # Instructions
    row = 4
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "Instructions: Document all evidence collected during assessment. Include screenshots, configuration exports, logs, reports, etc."
    ws[f"A{row}"].font = Font(italic=True, size=10)
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 35

    # Column Headers (Row 6)
    headers = [
        ("A", "Evidence ID", 15),
        ("B", "Evidence Type", 20),
        ("C", "Description", 40),
        ("D", "Related Requirement", 30),
        ("E", "Source Assessment", 25),
        ("F", "Date Collected", 16),
        ("G", "Collected By", 20),
        ("H", "Location/Link", 35),
        ("I", "Verification Status", 18),
        ("J", "Verified By", 20),
        ("K", "Verification Date", 16),
        ("L", "Notes", 35),
    ]

    row = 6
    for col, header, width in headers:
        cell = ws[f"{col}{row}"]
        cell.value = header
        apply_style(cell, styles["column_header"])
        ws.column_dimensions[col].width = width

    # Example row (Row 7)
    example_data = [
        "EV-001",
        "Screenshot",
        "Splunk SIEM dashboard",
        "Monitoring Platform Capabilities",
        "IMP-A.8.16.1 - Sheet 1",
        "07.01.2025",
        "J. Smith",
        "\\\\fileserver\\evidence\\splunk_dashboard.png",
        "Verified",
        "M. Jones",
        "08.01.2025",
        "Shows real-time alerting enabled"
    ]
    
    row = 7
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        apply_style(cell, styles["example_row"])

    # Data Entry Rows (Rows 8-107: 100 rows)
    validations = create_base_validations(ws)
    
    evidence_types_validation = DataValidation(
        type="list",
        formula1='"Screenshot,Configuration File,Log Export,Report,Policy Document,Email,Meeting Notes,Other"',
        allow_blank=False
    )
    ws.add_data_validation(evidence_types_validation)
    
    verification_status_validation = DataValidation(
        type="list",
        formula1='"Verified,Pending,Not Verified"',
        allow_blank=False
    )
    ws.add_data_validation(verification_status_validation)
    
    for data_row in range(8, 108):
        # Auto-generate Evidence ID
        ws[f"A{data_row}"] = f'="EV-"&TEXT(ROW()-7,"000")'
        
        for col_idx in range(2, 13):  # B to L
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            # Apply dropdowns
            col_letter = get_column_letter(col_idx)
            if col_letter == 'B':  # Evidence Type
                evidence_types_validation.add(cell)
            elif col_letter == 'I':  # Verification Status
                verification_status_validation.add(cell)
        
        ws.row_dimensions[data_row].height = 25

    # Evidence Summary (Starting Row 110)
    row = 110
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "EVIDENCE SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Total Evidence Items:"
    ws[f"B{row}"] = "=COUNTA(A8:A107)"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].font = Font(bold=True)

    row += 1
    ws[f"A{row}"] = "Verified Evidence:"
    ws[f"B{row}"] = '=COUNTIF(I8:I107,"Verified")'
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].font = Font(bold=True)

    row += 1
    ws[f"A{row}"] = "Pending Verification:"
    ws[f"B{row}"] = '=COUNTIF(I8:I107,"Pending")'
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].font = Font(bold=True)

    row += 1
    ws[f"A{row}"] = "Not Verified:"
    ws[f"B{row}"] = '=COUNTIF(I8:I107,"Not Verified")'
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].font = Font(bold=True)


# ============================================================================
# SECTION 11: SHEET 9 - APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff_sheet(ws, styles):
    """Create Approval Sign-Off sheet for governance."""
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "APPROVAL SIGN-OFF"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    ws["A2"] = "Monitoring Infrastructure Assessment Approval Workflow"
    apply_style(ws["A2"], styles["subheader"])
    ws.row_dimensions[2].height = 25

    # Document Summary
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "DOCUMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    doc_summary = [
        ("Document ID:", "ISMS-IMP-A.8.16.1"),
        ("Assessment Area:", "Monitoring Infrastructure"),
        ("Assessment Date:", "[From Instructions sheet]"),
        ("Overall Compliance Score:", "[From Summary Dashboard]"),
        ("Critical Gaps Identified:", "[Count from Summary Dashboard]"),
        ("Assessor:", "[From Instructions sheet]"),
    ]

    row += 1
    for label, value in doc_summary:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:F{row}")
        row += 1

    # Assessment Prepared By
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "ASSESSMENT PREPARED BY"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    prepared_fields = [
        ("Name:", ""),
        ("Title:", ""),
        ("Date:", ""),
        ("Signature:", ""),
    ]

    row += 1
    for label, value in prepared_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.merge_cells(f"B{row}:F{row}")
        ws.row_dimensions[row].height = 25
        row += 1

    # Reviewed By (SOC Lead)
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "REVIEWED BY (SOC LEAD)"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for label, value in prepared_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.merge_cells(f"B{row}:F{row}")
        ws.row_dimensions[row].height = 25
        row += 1

    # Review Comments
    row += 1
    ws[f"A{row}"] = "Review Comments:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 1
    ws.merge_cells(f"A{row}:F{row+3}")
    ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    for i in range(4):
        ws.row_dimensions[row + i].height = 25
    row += 4

    # Reviewed By (Security Engineering)
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "REVIEWED BY (SECURITY ENGINEERING)"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for label, value in prepared_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.merge_cells(f"B{row}:F{row}")
        ws.row_dimensions[row].height = 25
        row += 1

    # Review Comments
    row += 1
    ws[f"A{row}"] = "Review Comments:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 1
    ws.merge_cells(f"A{row}:F{row+3}")
    ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    for i in range(4):
        ws.row_dimensions[row + i].height = 25
    row += 4

    # Approved By (CISO)
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    approval_fields = [
        ("Name:", ""),
        ("Title:", ""),
        ("Approval Decision:", "[Dropdown: Approved, Approved with Conditions, Rejected]"),
        ("Date:", ""),
        ("Signature:", ""),
    ]

    for label, value in approval_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.merge_cells(f"B{row}:F{row}")
        ws.row_dimensions[row].height = 25
        
        # Add dropdown for Approval Decision
        if "Approval Decision" in label:
            approval_validation = DataValidation(
                type="list",
                formula1='"Approved,Approved with Conditions,Rejected"',
                allow_blank=False
            )
            ws.add_data_validation(approval_validation)
            approval_validation.add(ws[f"B{row}"])
        
        row += 1

    # Conditions/Comments
    row += 1
    ws[f"A{row}"] = "Conditions/Comments:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 1
    ws.merge_cells(f"A{row}:F{row+4}")
    ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    for i in range(5):
        ws.row_dimensions[row + i].height = 25
    row += 5

    # Executive Acknowledgment (Optional)
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "EXECUTIVE ACKNOWLEDGMENT (OPTIONAL)"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    exec_fields = [
        ("Name:", ""),
        ("Title:", ""),
        ("Date:", ""),
        ("Signature:", ""),
    ]

    for label, value in exec_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.merge_cells(f"B{row}:F{row}")
        ws.row_dimensions[row].height = 25
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 25


# ============================================================================
# SECTION 12: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Philosophy: "The first principle is that you must not fool yourself
    — and you are the easiest person to fool." - Richard Feynman
    
    This is Systems Engineering applied to ISMS compliance:
    We assess CAPABILITIES, not checkboxes. We gather EVIDENCE, not platitudes.
    We implement DETECTIVE CONTROLS, not cargo cult rituals.
    """
    print("=" * 78)
    print("ISMS-IMP-A.8.16.1 - Monitoring Infrastructure Assessment Generator")
    print("ISO/IEC 27001:2022 Control A.8.16: Monitoring Activities")
    print("=" * 78)
    print("\n🎯 Systems Engineering Approach: Evidence-Based Monitoring Assessment")
    print("📊 Comprehensive: Platform, Coverage, Architecture, Integration, Performance")
    print("🔒 Audit-Ready: 100 evidence entries, 75 compliance checkpoints")
    print("\n" + "─" * 78)

    # Create workbook and setup styles
    print("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    print("\u2705 Workbook created with 9 sheets")

    # Create all sheets
    print("\n[Phase 2] Generating assessment sheets...")
    
    print("  [1/9] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    print("  \u2705 Instructions complete")

    print("  [2/9] Creating Monitoring Platform Capabilities...")
    create_monitoring_platform_sheet(wb["1. Monitoring Platform"], styles)
    print("  \u2705 Platform assessment complete (15 capability checks)")

    print("  [3/9] Creating Log Source Coverage...")
    create_log_source_coverage_sheet(wb["2. Log Source Coverage"], styles)
    print("  \u2705 Coverage assessment complete (33 log source rows, 15 checks)")

    print("  [4/9] Creating Data Collection Architecture...")
    create_data_collection_architecture_sheet(wb["3. Data Collection Arch"], styles)
    print("  \u2705 Architecture assessment complete (18 component rows, 15 checks)")

    print("  [5/9] Creating Integration & Enrichment...")
    create_integration_enrichment_sheet(wb["4. Integration Enrichment"], styles)
    print("  \u2705 Integration assessment complete (15 integration rows, 15 checks)")

    print("  [6/9] Creating Performance & Scalability...")
    create_performance_scalability_sheet(wb["5. Performance Scale"], styles)
    print("  \u2705 Performance assessment complete (15 metrics, 10 scalability factors)")

    print("  [7/9] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)
    print("  \u2705 Dashboard complete (consolidated compliance view)")

    print("  [8/9] Creating Evidence Register...")
    create_evidence_register_sheet(wb["Evidence Register"], styles)
    print("  \u2705 Evidence register complete (100 evidence rows)")

    print("  [9/9] Creating Approval Sign-Off...")
    create_approval_signoff_sheet(wb["Approval Sign-Off"], styles)
    print("  \u2705 Approval workflow complete (4-level sign-off)")

    # Save workbook
    print("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.16.1_Monitoring_Infrastructure_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        wb.save(filename)
        print(f"\u2705 SUCCESS: {filename}")
    except Exception as e:
        print(f"\u274C ERROR saving workbook: {e}")
        return 1

    # Summary
    print("\n" + "=" * 78)
    print("\u1F4CB WORKBOOK STRUCTURE SUMMARY")
    print("=" * 78)
    print("\n📄 Assessment Sheets:")
    print("  \u2022 Instructions & Legend (usage guidance, definitions)")
    print("  \u2022 1. Monitoring Platform Capabilities (SIEM/IDS/EDR/NDR assessment)")
    print("  \u2022 2. Log Source Coverage (33 system rows, coverage tracking)")
    print("  \u2022 3. Data Collection Architecture (18 component rows, security)")
    print("  \u2022 4. Integration & Enrichment (15 integration rows, threat intel)")
    print("  \u2022 5. Performance & Scalability (15 metrics, capacity planning)")
    print("\n📊 Consolidation & Governance:")
    print("  \u2022 Summary Dashboard (overall compliance, key findings)")
    print("  \u2022 Evidence Register (100 evidence entries)")
    print("  \u2022 Approval Sign-Off (4-level approval workflow)")
    print("\n" + "─" * 78)
    print("📈 ASSESSMENT CAPABILITIES:")
    print("  \u2022 75 compliance checkpoint items across 5 assessment areas")
    print("  \u2022 33 log source tracking rows")
    print("  \u2022 18 collection architecture component rows")
    print("  \u2022 15 integration/enrichment tracking rows")
    print("  \u2022 15 performance metrics + 10 scalability factors")
    print("  \u2022 100 evidence documentation entries")
    print("  \u2022 Automated compliance % calculations")
    print("  \u2022 Gap identification and priority tracking")
    print("\n" + "─" * 78)
    print("🎯 KEY FEATURES:")
    print("  \u2705 Platform-agnostic (works with ANY SIEM/monitoring solution)")
    print("  \u2705 Comprehensive evidence collection")
    print("  \u2705 Automated compliance calculations")
    print("  \u2705 Coverage assessment (critical systems priority)")
    print("  \u2705 Architecture security validation")
    print("  \u2705 Integration health tracking")
    print("  \u2705 Performance & capacity monitoring")
    print("  \u2705 Multi-level approval workflow")
    print("  \u2705 Semi-annual review cycle support")
    print("\n" + "=" * 78)
    print("🚀 NEXT STEPS:")
    print("  1. Open the generated workbook")
    print("  2. Complete Instructions & Legend sheet first")
    print("  3. Fill in YOUR monitoring platforms (vendor-agnostic approach)")
    print("  4. Document ALL log sources (prioritize Critical/High systems)")
    print("  5. Assess collection architecture security")
    print("  6. Evaluate integrations and enrichment")
    print("  7. Review performance metrics and capacity")
    print("  8. Check Summary Dashboard for gaps")
    print("  9. Document evidence in Evidence Register")
    print("  10. Obtain final approval via Approval Sign-Off")
    print("\n💡 PRO TIP:")
    print("  This workbook is technology-independent. Whether you use Splunk,")
    print("  Elastic, QRadar, Sentinel, or ANY other SIEM - this framework")
    print("  assesses MONITORING CAPABILITIES, not brand names.")
    print("  That's Systems Engineering. That's Feynman-approved ISMS.")
    print("\n" + "=" * 78)
    print('\n"The first principle is that you must not fool yourself')
    print('— and you are the easiest person to fool." - Richard Feynman')
    print("\n🎓 This is not cargo cult ISMS. This is evidence-based monitoring.")
    print("🔍 We assess DETECTIVE CONTROLS, not checkbox compliance.")
    print("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    exit(main())