#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.16.5 - Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.16: Monitoring Activities
Assessment Domain 5 of 5: Consolidated Compliance Dashboard (MASTER WORKBOOK)

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific dashboard requirements, KPI definitions, and
reporting structure.

Key customization areas:
1. External workbook paths (update to your actual file locations)
2. KPI definitions and thresholds (per your governance requirements)
3. Compliance scoring criteria (aligned with your audit framework)
4. Trend analysis periods (based on your reporting cycles)
5. Approval workflow (per your organizational structure)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for monitoring)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates the MASTER Excel dashboard that consolidates data from
all four A.8.16 monitoring assessment workbooks into a single executive view
for compliance oversight and audit readiness.

**Purpose:**
Provides executive management and auditors with consolidated A.8.16 compliance
status, key performance indicators (KPIs), gap analysis, and trend tracking
across all monitoring assessment domains.

**Dashboard Consolidates:**
- ISMS-IMP-A.8.16.1: Monitoring Infrastructure Assessment
- ISMS-IMP-A.8.16.2: Baseline & Anomaly Detection Assessment
- ISMS-IMP-A.8.16.3: Coverage Assessment
- ISMS-IMP-A.8.16.4: Alert Management & Response Assessment

**Generated Dashboard Structure:**
1. Instructions & Legend - Dashboard usage and data source information
2. Executive Summary - High-level compliance status and critical findings
3. Compliance Matrix - Detailed compliance by assessment domain
4. KPIs - Key Performance Indicators (MTTD, MTTR, Coverage, Detection Rate)
5. Gap Remediation Tracker - Consolidated gaps with priorities and timelines
6. Trend Analysis - Historical compliance and performance trends
7. Evidence & Approvals - Evidence completeness and approval workflow

**Key Features:**
- Automated data consolidation via Excel external workbook references
- Real-time compliance calculations across all assessment domains
- KPI visualization with conditional formatting
- Gap prioritization and remediation tracking
- Trend analysis with time-series metrics
- Evidence completeness validation
- Multi-stakeholder approval workflow
- Audit-ready format and presentation

**CRITICAL PREREQUISITE:**
This dashboard requires NORMALIZED assessment workbooks. Run the normalization
script BEFORE generating this dashboard:

    python3 normalize_assessment_files_a816.py

Dashboard will show #REF errors if normalized source files are missing or
incorrectly named.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a816_5_compliance_dashboard.py

Output:
    File: ISMS_A_8_16_5_Compliance_Dashboard_YYYYMMDD.xlsx
    Location: Dashboard_Sources/ directory (place with normalized workbooks)

Setup Process (CRITICAL):
    1. Generate all four assessment workbooks (Scripts 1-4)
    2. Complete assessments (populate data in workbooks)
    3. Run normalization script:
       python3 normalize_assessment_files_a816.py
    4. Verify normalized files exist in Dashboard_Sources/:
       - A816_1_Monitoring_Infrastructure.xlsx
       - A816_2_Baseline_Detection.xlsx
       - A816_3_Coverage_Assessment.xlsx
       - A816_4_Alert_Management.xlsx
    5. Generate dashboard (this script)
    6. Place dashboard in Dashboard_Sources/ folder
    7. Open dashboard in Excel
    8. Click "Update Links" when prompted
    9. Review consolidated compliance data
    10. Complete manual sections (Gap Remediation, Trend Analysis)
    11. Obtain multi-level approvals

Post-Generation Steps:
    1. Review Executive Summary for critical findings
    2. Validate Compliance Matrix accuracy
    3. Analyze KPI trends (MTTD, MTTR, Coverage)
    4. Prioritize gaps in Gap Remediation Tracker
    5. Update Trend Analysis with historical data
    6. Verify evidence completeness
    7. Obtain stakeholder approvals
    8. Present to CISO and executive management
    9. Provide to auditors as primary A.8.16 compliance evidence

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.16
Assessment Domain:    5 of 5 (Compliance Dashboard - MASTER)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organization License/Terms]

Related Documents:
    - ISMS-POL-A.8.16: Monitoring Activities Policy (Governance)
    - ISMS-POL-A.8.16 (All Sections): Policy framework
    - ISMS-IMP-A.8.16.1: Monitoring Infrastructure Assessment
    - ISMS-IMP-A.8.16.2: Baseline & Detection Assessment
    - ISMS-IMP-A.8.16.3: Coverage Assessment
    - ISMS-IMP-A.8.16.4: Alert Management Assessment
    - ISMS-IMP-A.8.16.5: Compliance Dashboard Specification

Supporting Scripts:
    - normalize_assessment_files_a816.py (REQUIRED PREREQUISITE)
    - consolidate_a816_dashboard.py (Alternative consolidation method)
    - excel_sanity_check_a816.py (Quality assurance validation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 24.01.2025
    - Initial release
    - Implements consolidated dashboard per ISMS-IMP-A.8.16.5 specification
    - Integrates all four A.8.16 assessment domains
    - Provides executive-level compliance overview
    - Includes KPI tracking and trend analysis
    - Audit-ready format

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Dashboard Philosophy:**
"The first principle is that you must not fool yourself — and you are the 
easiest person to fool." - Richard Feynman

This dashboard consolidates objective compliance metrics from evidence-based
assessments. It is not a self-assessment checklist or compliance theater.
Every metric has source data in the underlying assessment workbooks.

**Critical Success Factors:**

1. **Complete Underlying Assessments First**
   Dashboard is only as good as source data. Garbage in = garbage out.

2. **Normalize Before Consolidating**
   Run normalize_assessment_files_a816.py BEFORE generating dashboard.
   Skip this and dashboard will be full of #REF errors.

3. **Place in Correct Directory**
   Dashboard MUST be in same folder as normalized source workbooks:
   Dashboard_Sources/ISMS_A_8_16_5_Compliance_Dashboard_YYYYMMDD.xlsx

4. **Update External Links**
   Excel will prompt to update links when opening. Click "Update" to refresh
   data from source workbooks.

**Overall Compliance Calculation:**

Overall A.8.16 Compliance Score = Average of:
- Monitoring Infrastructure Compliance (Domain 1)
- Baseline & Detection Compliance (Domain 2)
- Coverage Assessment Compliance (Domain 3)
- Alert Management Compliance (Domain 4)

Target: ≥90% overall compliance for mature monitoring program.

**Key Performance Indicators (KPIs):**

Tracked KPIs (per ISMS-POL-A.8.16):
- Mean Time to Detect (MTTD): Target ≤15 min (Critical), ≤1 hr (High)
- Mean Time to Respond (MTTR): Target ≤1 hr (Critical), ≤4 hr (High)
- Monitoring Coverage: Target 100% (Tier 1), 90% (Tier 2), 70% (Tier 3)
- Detection Effectiveness: Target ≥95% true positive rate
- False Positive Rate: Target <5%
- Alert Volume: Target <100 alerts/day (mature SOC)

**Gap Prioritization:**

Critical Gaps (Immediate Action Required):
- Tier 1 systems with <100% monitoring coverage
- MTTD/MTTR exceeding SLA for critical events
- Detection effectiveness <80%
- False positive rate >20% (alert fatigue risk)

High Priority Gaps:
- Tier 2 systems with <90% coverage
- Missing baselines for critical systems
- Escalation procedures not documented
- Evidence gaps for audit

**Audit Considerations:**

This dashboard is PRIMARY evidence for A.8.16 compliance in ISO 27001 audits.
Auditors will expect:
- Complete data from all four assessment domains
- Documented gaps with remediation plans
- KPI tracking with trend analysis
- Evidence of continuous improvement
- Executive oversight (approvals)

Incomplete dashboard = audit finding. Ensure all source assessments are
complete before presenting to auditors.

**Data Protection:**

Dashboard consolidates sensitive operational data including:
- Complete monitoring infrastructure details
- Coverage gaps and blind spots
- Performance metrics (including weaknesses)
- Remediation plans and priorities

Classification: INTERNAL - CONFIDENTIAL
Handle per organization's data classification policies.

**Common Dashboard Issues:**

1. **#REF Errors**
   - Cause: Normalized source workbooks missing or incorrectly named
   - Fix: Run normalize_assessment_files_a816.py and verify file names

2. **Zeros in KPIs**
   - Cause: Source assessments not completed
   - Fix: Complete all four assessment workbooks first

3. **Update Links Prompt**
   - Cause: Normal Excel behavior for external references
   - Action: Click "Update" to refresh from source workbooks

4. **Formula Errors After Moving Files**
   - Cause: External workbook references use relative paths
   - Fix: Keep dashboard and source workbooks in same folder

**Maintenance:**

Review and update dashboard:
- Monthly: Update KPI metrics and gap status
- Quarterly: Complete compliance review with executive management
- Semi-annually: Trend analysis update
- Annually: Full A.8.16 reassessment with updated dashboard
- Ad-hoc: After major infrastructure changes or incidents

**Quality Assurance:**

Run sanity check before distributing:
    python3 excel_sanity_check_a816.py ISMS_A_8_16_5_*.xlsx

Validate dashboard with SOC management and CISO before presenting to auditors
or executive management.

**Executive Presentation:**

Dashboard is designed for executive consumption:
- Executive Summary: High-level status (5-minute brief)
- KPIs: Performance metrics and trends
- Gap Remediation: Action items and priorities
- Evidence & Approvals: Sign-off section

Use color-coded status (Green/Yellow/Red) for executive visibility.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys

# =============================================================================
# Third-Party Imports
# =============================================================================
try:
    import openpyxl
except ImportError:
    logger.error("\u274C Error: openpyxl not installed")
    logger.info("ℹ️  Install with: sudo apt install python3-openpyxl")
    logger.info("   or: pip3 install openpyxl")
    sys.exit(1)


# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.16.5"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.8.16"
CONTROL_NAME = "Monitoring Activities"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches markdown specification
    sheets = [
        "Instructions & Legend",
        "Executive Summary",
        "Compliance Matrix",
        "KPIs",
        "Gap Remediation Tracker",
        "Trend Analysis",
        "Evidence & Approvals",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "critical_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "formula_cell": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_green": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_yellow": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_red": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# ============================================================================
# SECTION 2: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_legend(ws, styles):
    """Create Instructions & Legend sheet."""
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.8.16.5 – Monitoring Activities Compliance Dashboard\n"
        "ISO/IEC 27001:2022 - Control A.8.16: Monitoring Activities - Master Compliance Overview"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 50

    # Document Information
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.8.16.5"),
        ("Report Type", "Compliance Summary Dashboard"),
        ("Related Policy", "ISMS-POL-A.8.16 (Monitoring Activities)"),
        ("Version", "1.0"),
        ("Report Date", ""),
        ("Reporting Period", ""),
        ("Prepared By", ""),
        ("Organization", ""),
        ("Review Cycle", "Quarterly"),
        ("Last Updated", "=TODAY()"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        elif "=" in str(value):
            ws[f"B{row}"].font = Font(color="0000FF")
        row += 1

    # Purpose
    row += 2
    ws[f"A{row}"] = "Purpose"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = (
        "This dashboard aggregates compliance data from four assessment workbooks to provide "
        "executive oversight of the organization's monitoring activities program per ISO/IEC 27001:2022 Control A.8.16."
    )
    ws[f"A{row}"].alignment = Alignment(wrap_text=True)

    # Source Workbooks
    row += 3
    ws[f"A{row}"] = "Source Assessment Workbooks"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    
    sources = [
        ("ISMS-IMP-A.8.16.1", "Monitoring Infrastructure Assessment", "Log sources, SIEM, tools"),
        ("ISMS-IMP-A.8.16.2", "Baseline & Detection Assessment", "Baselines, detection rules"),
        ("ISMS-IMP-A.8.16.3", "Coverage Assessment", "Asset coverage, monitoring scope"),
        ("ISMS-IMP-A.8.16.4", "Alert Management Assessment", "Alerting, response, escalation"),
    ]
    
    row += 1
    for doc_id, title, scope in sources:
        ws[f"A{row}"] = doc_id
        ws[f"B{row}"] = title
        ws[f"C{row}"] = scope
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

    # How to Use
    row += 2
    ws[f"A{row}"] = "How to Use This Dashboard"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    
    instructions = [
        "1. PREREQUISITE: Run normalization script first:",
        "   python3 normalize_assessment_files_a816.py",
        "",
        "2. Place this dashboard in the Dashboard_Sources folder with normalized files",
        "",
        "3. Open dashboard and click 'Update Links' when prompted",
        "",
        "4. Executive Summary will auto-populate with current compliance data",
        "",
        "5. Review compliance metrics and gaps across all assessment areas",
        "",
        "6. Complete manual sections (Gap Remediation, Trend Analysis, Evidence)",
        "",
        "7. Obtain approvals via Evidence & Approvals sheet",
    ]
    
    row += 1
    for instruction in instructions:
        ws[f"A{row}"] = instruction
        if instruction.startswith("   "):
            ws[f"A{row}"].font = Font(italic=True, color="666666")
        row += 1

    # Legend
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "LEGEND"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    legend_headers = ["Symbol", "Status", "Meaning", "When to Use"]
    row += 1
    for col_idx, header in enumerate(legend_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    legend_data = [
        ("\u2705", "Compliant", "Requirement fully met", "All controls implemented and effective"),
        ("\u26A0\uFE0F", "Partial", "Requirement partially met", "Some controls in place, gaps remain"),
        ("\u274C", "Non-Compliant", "Requirement not met", "Controls absent or ineffective"),
        ("N/A", "Not Applicable", "Requirement does not apply", "Requirement not relevant to environment"),
    ]

    row += 1
    for symbol, status, meaning, when in legend_data:
        ws[f"A{row}"] = symbol
        ws[f"B{row}"] = status
        ws[f"C{row}"] = meaning
        ws[f"D{row}"] = when
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = styles["border"]
        row += 1

    # Color Code Legend
    row += 2
    ws[f"A{row}"] = "Color Coding"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    
    row += 1
    ws[f"A{row}"] = "Yellow cells"
    ws[f"A{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"] = "User input required"
    
    row += 1
    ws[f"A{row}"] = "Gray cells"
    ws[f"A{row}"].fill = styles["formula_cell"]["fill"]
    ws[f"B{row}"] = "Auto-calculated (formulas)"
    
    row += 1
    ws[f"A{row}"] = "Green cells"
    ws[f"A{row}"].fill = styles["status_green"]["fill"]
    ws[f"B{row}"] = "Compliant status"
    
    row += 1
    ws[f"A{row}"] = "Yellow cells"
    ws[f"A{row}"].fill = styles["status_yellow"]["fill"]
    ws[f"B{row}"] = "Partial/Warning status"
    
    row += 1
    ws[f"A{row}"] = "Red cells"
    ws[f"A{row}"].fill = styles["status_red"]["fill"]
    ws[f"B{row}"] = "Non-compliant/Critical status"

    # Important Notes
    row += 3
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "IMPORTANT NOTES"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    notes = [
        "\u2022 External Links: This dashboard uses Excel external links to source workbooks",
        "\u2022 File Location: Keep dashboard and source files in the same directory",
        "\u2022 Update Links: Always click 'Update Links' when opening the dashboard",
        "\u2022 Data Integrity: Do not modify source workbook filenames after normalization",
        "\u2022 Backup: Maintain backups of source assessment workbooks",
        "\u2022 Version Control: Track dashboard versions when compliance data changes",
    ]

    row += 1
    for note in notes:
        ws[f"A{row}"] = note
        ws[f"A{row}"].font = Font(color="C00000")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.freeze_panes = "A4"


# ============================================================================
# END OF PART 1
# ============================================================================

# ============================================================================
# SECTION 3: EXECUTIVE SUMMARY WITH EXTERNAL WORKBOOK LINKS
# ============================================================================

def create_executive_summary(ws, styles):
    """Create Executive Summary with external workbook references."""
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:H1")
    ws["A1"] = (
        "MONITORING ACTIVITIES - EXECUTIVE SUMMARY\n"
        "ISO/IEC 27001:2022 Control A.8.16 Compliance Status"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 50

    # ---------- DOCUMENT INFORMATION ----------
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Dashboard ID", "ISMS-IMP-A.8.16.5"),
        ("Assessment Date", ""),
        ("Reporting Period", ""),
        ("Next Review", ""),
        ("Prepared By", ""),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    # ---------- OVERALL COMPLIANCE STATUS ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "OVERALL MONITORING COMPLIANCE STATUS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    # Compliance Scorecard
    row += 1
    scorecard_headers = ["Metric", "Score", "Target", "Status"]
    for col_idx, header in enumerate(scorecard_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    scorecard_start = row + 1
    scorecard_metrics = [
        ("Overall Compliance Rate", "=AVERAGE(C17:C20)", "≥95%", ""),
        ("Critical Gaps", "=COUNTIFS('Compliance Matrix'!I:I,\"Critical\",'Compliance Matrix'!F:F,\"\u274C Non-Compliant\")", "0", ""),
        ("High-Risk Items", "=COUNTIFS('Gap Remediation Tracker'!E:E,\"High\",'Gap Remediation Tracker'!K:K,\"<>Resolved\")", "≤5", ""),
        ("Remediation Progress", "='Gap Remediation Tracker'!L3", "≥80%", ""),
    ]

    row += 1
    for metric, formula, target, status in scorecard_metrics:
        ws.cell(row=row, column=1, value=metric).border = styles["border"]
        ws.cell(row=row, column=2, value=formula).font = Font(bold=True, color="0000FF")
        ws.cell(row=row, column=2).border = styles["border"]
        ws.cell(row=row, column=3, value=target).border = styles["border"]
        ws.cell(row=row, column=4, value=status).border = styles["border"]
        row += 1

    # ---------- COMPLIANCE BY ASSESSMENT AREA (WITH EXTERNAL LINKS) ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "COMPLIANCE BY ASSESSMENT AREA"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    area_headers = ["Assessment", "Domain", "Compliance %", "Status", "Trend"]
    row += 1
    for col_idx, header in enumerate(area_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # CRITICAL: External workbook references pulling data from normalized source files
    # These formulas read from the "Summary Dashboard" sheet, row 13 (OVERALL TOTAL)
    # Calculating compliance % from: Compliant/(Total-N/A)*100
    # Row 13 because: rows 7-11 have 5 areas, then blank row 12, then OVERALL at row 13
    # Using direct calculation instead of column G (which contains text "XX.X%")
    
    assessment_areas = [
        ("IMP-A.8.16.1", "Infrastructure", 
         "=[ISMS-IMP-A.8.16.1.xlsx]'Summary Dashboard'!$B$13/($F$13-$E$13)*100", 
         "=IF(C17>=95,\"\u2705 Compliant\",IF(C17>=80,\"\u26A0\uFE0F Partial\",\"\u274C Non-Compliant\"))", 
         "→"),
        ("IMP-A.8.16.2", "Baseline & Detection", 
         "=[ISMS-IMP-A.8.16.2.xlsx]'Summary Dashboard'!$B$13/($F$13-$E$13)*100",
         "=IF(C18>=95,\"\u2705 Compliant\",IF(C18>=80,\"\u26A0\uFE0F Partial\",\"\u274C Non-Compliant\"))", 
         "→"),
        ("IMP-A.8.16.3", "Coverage", 
         "=[ISMS-IMP-A.8.16.3.xlsx]'Summary Dashboard'!$B$13/($F$13-$E$13)*100",
         "=IF(C19>=95,\"\u2705 Compliant\",IF(C19>=80,\"\u26A0\uFE0F Partial\",\"\u274C Non-Compliant\"))", 
         "↑"),
        ("IMP-A.8.16.4", "Alert Management", 
         "=[ISMS-IMP-A.8.16.4.xlsx]'Summary Dashboard'!$B$13/($F$13-$E$13)*100",
         "=IF(C20>=95,\"\u2705 Compliant\",IF(C20>=80,\"\u26A0\uFE0F Partial\",\"\u274C Non-Compliant\"))", 
         "↑"),
    ]

    row += 1
    for assessment, domain, compliance_formula, status_formula, trend in assessment_areas:
        ws[f"A{row}"] = assessment
        ws[f"B{row}"] = domain
        ws[f"C{row}"] = compliance_formula
        ws[f"C{row}"].font = Font(bold=True, color="0000FF")
        ws[f"C{row}"].number_format = "0.0\"%\""
        ws[f"D{row}"] = status_formula
        ws[f"D{row}"].font = Font(bold=True)
        ws[f"E{row}"] = trend
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row}"].border = styles["border"]
        row += 1

    # OVERALL row (aggregates all 4 assessments)
    ws[f"A{row}"] = "OVERALL"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    ws[f"B{row}"] = "All Domains"
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"C{row}"] = "=AVERAGE(C17:C20)"
    ws[f"C{row}"].font = Font(bold=True, size=11, color="0000FF")
    ws[f"C{row}"].number_format = "0.0\"%\""
    ws[f"D{row}"] = "=IF(C21>=95,\"\u2705 Compliant\",IF(C21>=80,\"\u26A0\uFE0F Partial\",\"\u274C Non-Compliant\"))"
    ws[f"D{row}"].font = Font(bold=True, size=11)
    ws[f"E{row}"] = "→"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].border = styles["border"]
        ws[f"{col}{row}"].fill = PatternFill(start_color="D8E4F8", end_color="D8E4F8", fill_type="solid")

    # ---------- CRITICAL METRICS SUMMARY ----------
    row += 3
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "CRITICAL METRICS SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    metrics_headers = ["Metric", "Current", "Target", "Status"]
    row += 1
    for col_idx, header in enumerate(metrics_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Pull key metrics from source workbooks OR manual input
    # Note: Specific metric locations vary by workbook layout
    # Users can copy-paste these values from source workbook Summary Dashboards
    critical_metrics = [
        # Infrastructure Metrics (from IMP-A.8.16.1)
        ("COVERAGE METRICS", "", "", ""),
        ("Critical Systems Monitored", "", "100%", 
         "=IF(B28>=100,\"\u2705\",IF(B28>=80,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("Overall System Coverage", "", ">80%",
         "=IF(B29>=80,\"\u2705\",IF(B29>=60,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("Network Segment Coverage", "", "100%",
         "=IF(B30>=100,\"\u2705\",IF(B30>=80,\"\u26A0\uFE0F\",\"\u274C\"))"),
        
        # Detection Metrics (from IMP-A.8.16.2)
        ("", "", "", ""),
        ("DETECTION METRICS", "", "", ""),
        ("Baseline Coverage (Critical)", "", "100%",
         "=IF(B33>=100,\"\u2705\",IF(B33>=80,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("Detection Rate", "", ">90%",
         "=IF(B34>=90,\"\u2705\",IF(B34>=70,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("False Positive Rate", "", "<25%",
         "=IF(B35<=25,\"\u2705\",IF(B35<=40,\"\u26A0\uFE0F\",\"\u274C\"))"),
        
        # Response Metrics (from IMP-A.8.16.4)
        ("", "", "", ""),
        ("RESPONSE METRICS", "", "", ""),
        ("MTTD (Critical Alerts)", "", "<5 min",
         "=IF(B38<=5,\"\u2705\",IF(B38<=15,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("MTTR (Critical Incidents)", "", "<8 hrs",
         "=IF(B39<=8,\"\u2705\",IF(B39<=24,\"\u26A0\uFE0F\",\"\u274C\"))"),
        ("SLA Compliance", "", ">95%",
         "=IF(B40>=95,\"\u2705\",IF(B40>=85,\"\u26A0\uFE0F\",\"\u274C\"))"),
        
        # Operational Metrics (from IMP-A.8.16.1)
        ("", "", "", ""),
        ("OPERATIONAL METRICS", "", "", ""),
        ("Log Sources Integrated", "", "Target",
         "=IF(ISNUMBER(B43),IF(B43>=ISNUMBER(C43),\"\u2705\",\"\u26A0\uFE0F\"),\"\")"),
        ("Active Detection Rules", "", "Target",
         "=IF(ISNUMBER(B44),IF(B44>=ISNUMBER(C44),\"\u2705\",\"\u26A0\uFE0F\"),\"\")"),
        ("SOC Analyst Coverage", "", "24/7",
         "=IF(B45=\"24/7\",\"\u2705\",\"\u274C\")"),
    ]

    row += 1
    for metric, current, target, status in critical_metrics:
        # Section headers (bold, no formulas)
        if metric.endswith("METRICS"):
            ws.merge_cells(f"A{row}:D{row}")
            ws[f"A{row}"] = metric
            ws[f"A{row}"].font = Font(bold=True, size=10)
            ws[f"A{row}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        # Empty rows
        elif metric == "":
            pass
        # Regular metrics
        else:
            ws[f"A{row}"] = metric
            ws[f"B{row}"] = current
            if current == "":
                # Empty cell = user input required
                ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            elif current.startswith("="):
                ws[f"B{row}"].font = Font(color="0000FF")
            ws[f"C{row}"] = target
            ws[f"D{row}"] = status
            if status.startswith("="):
                ws[f"D{row}"].font = Font(bold=True)
            for col in ["A", "B", "C", "D"]:
                ws[f"{col}{row}"].border = styles["border"]
        row += 1

    # ---------- TOP 5 STRENGTHS ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "TOP 5 STRENGTHS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    strength_headers = ["Rank", "Strength", "Evidence"]
    row += 1
    for col_idx, header in enumerate(strength_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for rank in range(1, 6):
        ws[f"A{row}"] = rank
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{row}"].fill = styles["input_cell"]["fill"]
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].border = styles["border"]
        row += 1

    # ---------- TOP 5 GAPS ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "TOP 5 CRITICAL GAPS"
    ws[f"A{row}"].font = styles["critical_header"]["font"]
    ws[f"A{row}"].fill = styles["critical_header"]["fill"]
    ws[f"A{row}"].alignment = styles["critical_header"]["alignment"]

    gap_headers = ["Priority", "Gap Description", "Risk Level", "Remediation Target"]
    row += 1
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for priority in ["Critical", "High", "High", "Medium", "Medium"]:
        ws[f"A{row}"] = priority
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{row}"].fill = styles["input_cell"]["fill"]
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = styles["border"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.freeze_panes = "A4"


# ============================================================================
# END OF PART 2
# ============================================================================

# ============================================================================
# SECTION 4: REMAINING SHEETS (SIMPLIFIED)
# ============================================================================

def create_compliance_matrix(ws, styles):
    """Create Compliance Matrix sheet."""
    
    # Header
    ws.merge_cells("A1:L1")
    ws["A1"] = "DETAILED COMPLIANCE MATRIX"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # Column Headers
    headers = [
        ("A", "Policy Reference", 25),
        ("B", "Requirement", 40),
        ("C", "Control Type", 18),
        ("D", "Assessment Sheet", 22),
        ("E", "Evidence Location", 30),
        ("F", "Implementation Status", 20),
        ("G", "Compliance Status", 20),
        ("H", "Gap Description", 35),
        ("I", "Risk Level", 15),
        ("J", "Remediation Owner", 20),
        ("K", "Target Date", 14),
        ("L", "Notes", 25),
    ]

    row = 3
    for col_letter, header, width in headers:
        ws[f"{col_letter}{row}"] = header
        ws[f"{col_letter}{row}"].font = styles["column_header"]["font"]
        ws[f"{col_letter}{row}"].fill = styles["column_header"]["fill"]
        ws[f"{col_letter}{row}"].alignment = styles["column_header"]["alignment"]
        ws[f"{col_letter}{row}"].border = styles["border"]
        ws.column_dimensions[col_letter].width = width

    # Data rows (105 requirements from policy sections)
    row = 4
    for req_num in range(1, 106):
        ws[f"A{row}"] = f"REQ-{req_num:03d}"
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]  # Implementation Status
        ws[f"G{row}"].fill = styles["formula_cell"]["fill"]  # Compliance Status (formula-driven)
        ws[f"H{row}"].fill = styles["input_cell"]["fill"]  # Gap Description
        ws[f"I{row}"].fill = styles["input_cell"]["fill"]  # Risk Level
        ws[f"J{row}"].fill = styles["input_cell"]["fill"]  # Remediation Owner
        ws[f"K{row}"].fill = styles["input_cell"]["fill"]  # Target Date
        ws[f"L{row}"].fill = styles["input_cell"]["fill"]  # Notes
        
        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            ws[f"{col_letter}{row}"].border = styles["border"]
        row += 1

    ws.freeze_panes = "A4"


def create_kpis(ws, styles):
    """Create KPIs sheet."""
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "KEY PERFORMANCE INDICATORS (KPIs)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # Section 1: Coverage KPIs
    row = 3
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "COVERAGE KPIs"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    kpi_headers = ["KPI", "Current", "Target", "Baseline", "Trend (3mo)", "Status"]
    row += 1
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    coverage_kpis = [
        "Critical Systems Monitored %",
        "High Priority Systems %",
        "Overall Asset Coverage %",
        "Network Segments Covered %",
        "Log Sources Integrated (Count)",
        "Log Volume (GB/day)",
        "Log Collection Reliability %",
    ]

    row += 1
    for kpi in coverage_kpis:
        ws[f"A{row}"] = kpi
        ws[f"A{row}"].border = styles["border"]
        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = styles["border"]
        row += 1

    # Section 2: Detection KPIs
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "DETECTION KPIs"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    row += 1
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    detection_kpis = [
        "Tier 1 Systems with Baselines %",
        "Tier 2 Systems with Baselines %",
        "Baseline Staleness (Avg Days)",
        "Active Detection Rules (Count)",
        "MITRE ATT&CK Coverage %",
        "Detection Rate (Testing) %",
        "False Positive Rate %",
        "Rules Requiring Tuning (Count)",
    ]

    row += 1
    for kpi in detection_kpis:
        ws[f"A{row}"] = kpi
        ws[f"A{row}"].border = styles["border"]
        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = styles["border"]
        row += 1

    # Section 3: Response KPIs
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "RESPONSE KPIs"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    row += 1
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    response_kpis = [
        "Total Alerts (30d)",
        "True Positive Rate %",
        "MTTA - Critical (minutes)",
        "MTTT - Critical (minutes)",
        "MTTI - Critical (hours)",
        "MTTR - Critical (hours)",
        "SLA Compliance Rate %",
        "Escalation Rate %",
    ]

    row += 1
    for kpi in response_kpis:
        ws[f"A{row}"] = kpi
        ws[f"A{row}"].border = styles["border"]
        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = styles["border"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15
    ws.freeze_panes = "A4"


def create_gap_remediation_tracker(ws, styles):
    """Create Gap Remediation Tracker sheet."""
    
    # Header
    ws.merge_cells("A1:M1")
    ws["A1"] = "GAP ANALYSIS & REMEDIATION TRACKER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # Gap Inventory
    headers = [
        ("A", "Gap ID", 12),
        ("B", "Source Assessment", 18),
        ("C", "Gap Category", 20),
        ("D", "Description", 35),
        ("E", "Risk", 12),
        ("F", "Impact", 15),
        ("G", "Remediation Plan", 30),
        ("H", "Owner", 20),
        ("I", "Target Date", 14),
        ("J", "Budget", 15),
        ("K", "Status", 15),
        ("L", "% Complete", 12),
        ("M", "Last Updated", 14),
    ]

    row = 3
    for col_letter, header, width in headers:
        ws[f"{col_letter}{row}"] = header
        ws[f"{col_letter}{row}"].font = styles["column_header"]["font"]
        ws[f"{col_letter}{row}"].fill = styles["column_header"]["fill"]
        ws[f"{col_letter}{row}"].alignment = styles["column_header"]["alignment"]
        ws[f"{col_letter}{row}"].border = styles["border"]
        ws.column_dimensions[col_letter].width = width

    # Data rows (60 gaps)
    row = 4
    for gap_num in range(1, 61):
        ws[f"A{row}"] = f"GAP-{gap_num:03d}"
        for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
            ws[f"{col_letter}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col_letter}{row}"].border = styles["border"]
        row += 1

    ws.freeze_panes = "A4"


def create_trend_analysis(ws, styles):
    """Create Trend Analysis sheet."""
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "TREND ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # Section 1: Compliance Trend (12 months)
    row = 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "MONTHLY COMPLIANCE TREND (12 MONTHS)"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    trend_headers = ["Month", "Overall %", "Infrastructure %", "Baseline %", "Coverage %", "Alert Mgmt %"]
    row += 1
    for col_idx, header in enumerate(trend_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    months = ["Jan 2025", "Feb 2025", "Mar 2025", "Apr 2025", "May 2025", "Jun 2025",
              "Jul 2025", "Aug 2025", "Sep 2025", "Oct 2025", "Nov 2025", "Dec 2025"]

    row += 1
    for month in months:
        ws[f"A{row}"] = month
        ws[f"A{row}"].border = styles["border"]
        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = styles["border"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 15
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 18
    ws.freeze_panes = "A5"


def create_evidence_approvals(ws, styles):
    """Create Evidence & Approvals sheet."""
    
    # Header
    ws.merge_cells("A1:L1")
    ws["A1"] = "EVIDENCE REGISTER & APPROVALS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # Evidence Register
    headers = [
        ("A", "Evidence ID", 15),
        ("B", "Evidence Type", 20),
        ("C", "Description", 35),
        ("D", "Related Requirement", 22),
        ("E", "Source Assessment", 20),
        ("F", "Date Collected", 16),
        ("G", "Collected By", 20),
        ("H", "Location/Link", 30),
        ("I", "Verification Status", 20),
        ("J", "Verified By", 20),
        ("K", "Verification Date", 16),
        ("L", "Notes", 25),
    ]

    row = 3
    for col_letter, header, width in headers:
        ws[f"{col_letter}{row}"] = header
        ws[f"{col_letter}{row}"].font = styles["column_header"]["font"]
        ws[f"{col_letter}{row}"].fill = styles["column_header"]["fill"]
        ws[f"{col_letter}{row}"].alignment = styles["column_header"]["alignment"]
        ws[f"{col_letter}{row}"].border = styles["border"]
        ws.column_dimensions[col_letter].width = width

    # Data rows (100 evidence entries)
    row = 4
    for evidence_num in range(1, 101):
        ws[f"A{row}"] = f"EVD-{evidence_num:03d}"
        for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            ws[f"{col_letter}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col_letter}{row}"].border = styles["border"]
        row += 1

    # Approval Workflow
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVAL WORKFLOW"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    # Prepared By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT PREPARED BY"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    prep_fields = ["Name", "Title", "Date", "Signature"]
    row += 1
    for field in prep_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Reviewed By (SOC Lead)
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (SOC LEAD)"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    review_fields = ["Name", "Title", "Review Comments", "Date"]
    row += 1
    for field in review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Approved By (CISO)
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    approval_fields = ["Name", "Title", "Approval Decision", "Conditions/Comments", "Date"]
    row += 1
    for field in approval_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Approval Decision dropdown
    decision_cell_row = row - 3
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{decision_cell_row}"])

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 5: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.8.16.5 - Compliance Summary Dashboard Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.16: Monitoring Activities")
    logger.info("=" * 80)

    wb = create_workbook()
    styles = setup_styles()

    logger.info("\n[1/7] Creating Instructions & Legend sheet...")
    create_instructions_legend(wb["Instructions & Legend"], styles)

    logger.info("[2/7] Creating Executive Summary sheet (with external links)...")
    create_executive_summary(wb["Executive Summary"], styles)

    logger.info("[3/7] Creating Compliance Matrix sheet...")
    create_compliance_matrix(wb["Compliance Matrix"], styles)

    logger.info("[4/7] Creating KPIs sheet...")
    create_kpis(wb["KPIs"], styles)

    logger.info("[5/7] Creating Gap Remediation Tracker sheet...")
    create_gap_remediation_tracker(wb["Gap Remediation Tracker"], styles)

    logger.info("[6/7] Creating Trend Analysis sheet...")
    create_trend_analysis(wb["Trend Analysis"], styles)

    logger.info("[7/7] Creating Evidence & Approvals sheet...")
    create_evidence_approvals(wb["Evidence & Approvals"], styles)

    filename = f"ISMS-IMP-A.8.16.5_Compliance_Dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    logger.info(f"\n\u2705 SUCCESS: {filename}")
    logger.info("\n" + "=" * 80)
    logger.info("COMPLIANCE SUMMARY DASHBOARD - COMPLETE")
    logger.info("=" * 80)
    logger.info("\nWorkbook Structure (7 Sheets):")
    logger.info("  1. Instructions & Legend - Setup and usage guide")
    logger.info("  2. Executive Summary - Overall compliance status with external links")
    logger.info("  3. Compliance Matrix - 105 requirement entries with compliance tracking")
    logger.info("  4. KPIs - Coverage, Detection, Response, and Operational metrics")
    logger.info("  5. Gap Remediation Tracker - 60 gap entries with remediation tracking")
    logger.info("  6. Trend Analysis - 12-month compliance trends")
    logger.info("  7. Evidence & Approvals - 100 evidence entries + approval workflow")
    
    logger.info("\nExternal Links Configuration:")
    logger.info("  \u2022 Dashboard references 4 normalized source workbooks:")
    logger.info("    - ISMS-IMP-A.8.16.1.xlsx (Monitoring Infrastructure)")
    logger.info("    - ISMS-IMP-A.8.16.2.xlsx (Baseline & Detection)")
    logger.info("    - ISMS-IMP-A.8.16.3.xlsx (Coverage Assessment)")
    logger.info("    - ISMS-IMP-A.8.16.4.xlsx (Alert Management)")
    
    logger.info("\nIMPORTANT SETUP STEPS:")
    logger.info("  1. Run normalization script FIRST:")
    logger.info("     python3 normalize_assessment_files_a816.py")
    logger.info("")
    logger.info("  2. Place this dashboard in the Dashboard_Sources folder")
    logger.info("     alongside the 4 normalized source workbooks")
    logger.info("")
    logger.info("  3. Open dashboard and click 'Update Links' when prompted")
    logger.info("")
    logger.info("  4. Executive Summary will auto-populate with compliance data!")
    logger.info("")
    logger.info("  5. Complete manual entry sections:")
    logger.info("     - Compliance Matrix (consolidated compliance status)")
    logger.info("     - KPIs (current metric values)")
    logger.info("     - Gap Remediation Tracker (remediation plans)")
    logger.info("     - Trend Analysis (historical data)")
    logger.info("     - Evidence & Approvals (supporting documentation)")
    logger.info("")
    logger.info("=" * 80)
    logger.info("\n🎯 ALL 5 MONITORING ASSESSMENT TOOLS COMPLETE! 🎯")
    logger.info("\nYou now have:")
    logger.info("  \u2705 4 Assessment workbooks (Infrastructure, Baseline, Coverage, Alert Mgmt)")
    logger.info("  \u2705 1 Normalization script (prepares files for dashboard)")
    logger.info("  \u2705 1 Consolidated dashboard (executive oversight)")
    logger.info("\nNo more cargo cult compliance - this is the real deal! 💪")
    logger.info("=" * 80 + "\n")


if __name__ == "__main__":
    main()


# ============================================================================
# END OF SCRIPT - COMPLETE COMPLIANCE SUMMARY DASHBOARD GENERATOR
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
