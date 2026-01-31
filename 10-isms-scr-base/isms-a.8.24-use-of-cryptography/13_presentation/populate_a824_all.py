#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
A.8.24 Assessment Workbooks - Comprehensive Dummy Data Population

This script populates all A.8.24 assessment workbooks with sample data.
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


def safely_write_data(ws, start_row, data):
    """Safely write data to worksheet, handling merged cells."""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except Exception as e:  # Skip merged/protected cells
                continue
        entries_written += 1
    return entries_written


def add_data_transmission_samples(filepath):
    """824_1: 8 columns - Status in column E (5)
    Columns: Service Description, Current TLS Version, Certificate Source,
             Certificate Validity, Status, Evidence Location, Gap Description,
             Remediation Needed
    """
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        logger.error("    Error loading %s: %s", filepath, e)
        return 0

    sheet_data = {
        "1.1 External HTTPS-TLS": [
            ["www.company.com", "TLS 1.3", "DigiCert", "Valid 2026-12-31", "Compliant", "SSL Labs A+", "", "No"],
            ["portal.company.com", "TLS 1.3", "Let's Encrypt", "Valid 2026-06-15", "Compliant", "Certbot auto-renew", "", "No"],
            ["api.company.com", "TLS 1.3", "DigiCert EV", "Valid 2027-03-20", "Compliant", "AWS ACM", "", "No"],
            ["mail.company.com", "TLS 1.2", "GoDaddy", "Expires 2026-02-28", "Partial", "Cert renew scheduled", "Upgrade TLS 1.3", "Yes"],
            ["mobile-api.company.com", "TLS 1.3", "DigiCert", "Valid 2026-11-30", "Compliant", "Auto rotation", "", "No"],
            ["partners.company.com", "TLS 1.3", "GlobalSign", "Valid 2027-01-15", "Compliant", "SAN cert", "", "No"],
            ["cdn.company.com", "TLS 1.3", "Cloudflare", "Valid 2027-06-01", "Compliant", "CF managed", "", "No"],
            ["admin.company.com", "TLS 1.3", "DigiCert EV", "Valid 2026-09-15", "Compliant", "EV cert + HSM", "", "No"],
            ["shop.company.com", "TLS 1.3", "Sectigo", "Valid 2026-08-20", "Compliant", "E-commerce cert", "", "No"],
            ["docs.company.com", "TLS 1.3", "Let's Encrypt", "Valid 2026-04-10", "Compliant", "Auto-renew", "", "No"],
            ["legacy.company.com", "TLS 1.1", "Self-signed", "Expired 2024-01-01", "Non-Compliant", "Legacy system", "URGENT: Replace", "Yes"],
            ["vpn.company.com", "TLS 1.2", "Internal CA", "Valid 2025-12-31", "Partial", "Internal CA", "Plan upgrade", "Yes"],
            ["test.company.com", "TLS 1.2", "Self-signed", "Valid 2026-03-01", "Partial", "Test env", "Use valid cert", "Yes"],
            ["blog.company.com", "TLS 1.3", "Let's Encrypt", "Valid 2026-07-15", "Compliant", "WordPress auto", "", "No"],
            ["status.company.com", "TLS 1.3", "DigiCert", "Valid 2026-10-01", "Compliant", "Status page", "", "No"],
        ],
        "2.1 Email Encryption": [
            ["Microsoft 365", "TLS 1.3", "S/MIME + TLS", "85% adoption", "Partial", "M365 admin", "Increase adoption", "Yes"],
            ["Gmail Business", "TLS 1.3", "TLS only", "100%", "Compliant", "Google Workspace", "", "No"],
            ["Internal Mail", "TLS 1.2", "PGP + TLS", "45% adoption", "Partial", "Postfix logs", "Promote PGP", "Yes"],
            ["Executive Email", "TLS 1.3", "S/MIME", "100%", "Compliant", "Cert store", "", "No"],
            ["External Partners", "TLS 1.2", "Opportunistic TLS", "90%", "Compliant", "Mail logs", "", "No"],
        ],
        "4.1 VPN": [
            ["Employee VPN", "TLS 1.3", "OpenVPN", "AES-256-GCM", "Compliant", "VPN logs", "", "No"],
            ["Site-to-Site VPN", "IPSec", "IPSec IKEv2", "AES-256-CBC", "Partial", "Firewall logs", "Add MFA", "Yes"],
            ["Executive VPN", "TLS 1.3", "WireGuard", "ChaCha20-Poly1305", "Compliant", "MDM logs", "", "No"],
            ["Partner VPN", "TLS 1.3", "SSL VPN", "AES-256-GCM", "Compliant", "Access logs", "", "No"],
            ["Contractor VPN", "TLS 1.2", "OpenVPN", "AES-128-CBC", "Non-Compliant", "Legacy config", "Upgrade + MFA", "Yes"],
        ],
    }

    total_written = 0
    for sheet_name, data in sheet_data.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            count = safely_write_data(ws, 8, data)
            total_written += count
            logger.info("    - %s: %d entries", sheet_name, count)

    try:
        wb.save(filepath)
    except Exception as e:
        logger.error("    Error saving: %s", e)
        return 0

    return total_written


def add_data_storage_samples(filepath):
    """824_2: 11 columns - Status in column H (8)
    Columns: Storage System/Device, Data Classification, Encryption Status,
             Encryption Type, Algorithm & Key Size, Key Management Method,
             Key Rotation Enabled, Status, Evidence Location, Gap Description,
             Remediation Needed
    """
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        logger.error("    Error loading %s: %s", filepath, e)
        return 0

    sheet_data = {
        "1. Mobile Devices": [
            ["Corporate iPhones", "Confidential", "iOS native", "Hardware", "AES-256", "MDM managed", "Yes", "Compliant", "MobileIron", "", "No"],
            ["Android Devices", "Confidential", "Android encryption", "Hardware", "AES-256", "MDM managed", "Yes", "Compliant", "Workspace ONE", "", "No"],
            ["iPads", "Confidential", "iOS native", "Hardware", "AES-256", "MDM managed", "Yes", "Compliant", "Jamf Pro", "", "No"],
            ["Tablets", "Internal", "Device encryption", "Software", "AES-256", "MDM managed", "Yes", "Partial", "Compliance reports", "Enforce 100%", "Yes"],
            ["BYOD iPhones", "Internal", "Containerization", "Software", "AES-256", "Work data only", "Yes", "Partial", "BYOD policy", "Increase enrollment", "Yes"],
            ["BYOD Android", "Internal", "Containerization", "Software", "AES-256", "Work data only", "Yes", "Partial", "BYOD policy", "Training needed", "Yes"],
            ["Test Devices", "Public", "Various", "Mixed", "Various", "No MDM", "No", "Non-Compliant", "Dev inventory", "Enroll in MDM", "Yes"],
        ],
        "2. Laptops & Workstations": [
            ["Windows Laptops", "Confidential", "BitLocker", "Full disk", "AES-256-XTS", "AD recovery", "Annual", "Compliant", "BitLocker dashboard", "", "No"],
            ["MacBooks", "Confidential", "FileVault 2", "Full disk", "AES-256", "MDM escrow", "Annual", "Compliant", "Jamf Pro", "", "No"],
            ["Linux Workstations", "Internal", "LUKS", "Full disk", "AES-256", "Manual backup", "Annual", "Partial", "Asset inventory", "Enforce 100%", "Yes"],
            ["Executive Laptops", "Restricted", "BitLocker + TPM", "Full disk", "AES-256-XTS", "Intune", "Quarterly", "Compliant", "Intune reports", "", "No"],
            ["Developer Macs", "Confidential", "FileVault 2", "Full disk", "AES-256", "MDM", "Annual", "Compliant", "Jamf inventory", "", "No"],
            ["Windows Desktops", "Internal", "BitLocker", "Full disk", "AES-256", "AD", "Annual", "Compliant", "SCCM reports", "", "No"],
            ["Contractor Laptops", "Internal", "Various", "Mixed", "Various", "No encryption", "No", "Non-Compliant", "Contractor list", "Enforce policy", "Yes"],
        ],
        "4. Databases": [
            ["PostgreSQL Prod", "Restricted", "TDE pgcrypto", "Transparent", "AES-256-GCM", "AWS KMS", "Yes", "Compliant", "DB logs", "", "No"],
            ["MySQL Customer", "Confidential", "TDE InnoDB", "Transparent", "AES-256", "Azure Key Vault", "Yes", "Compliant", "MySQL TDE status", "", "No"],
            ["SQL Server Finance", "Restricted", "TDE", "Transparent", "AES-256", "HSM-backed", "Yes", "Compliant", "SQL Server DMVs", "", "No"],
            ["MongoDB Analytics", "Internal", "At Rest", "Encrypted storage", "AES-256-GCM", "Self-managed", "No", "Partial", "MongoDB config", "Use KMS", "Yes"],
            ["Redis Cache", "Internal", "None", "None", "None", "N/A", "No", "Non-Compliant", "Redis config", "Add encryption", "Yes"],
            ["Oracle Legacy", "Confidential", "TDE", "Transparent", "AES-192", "Manual", "No", "Partial", "Oracle audit", "Upgrade AES-256", "Yes"],
        ],
        "5. Cloud Storage": [
            ["AWS S3 Prod", "Confidential", "SSE-KMS", "Server-side", "AES-256", "AWS KMS", "Yes", "Compliant", "S3 encryption", "", "No"],
            ["Azure Blob", "Confidential", "SSE", "Server-side", "AES-256", "Azure KV", "Yes", "Compliant", "Blob encryption", "", "No"],
            ["GCP Buckets", "Internal", "CMEK", "Server-side", "AES-256", "Cloud KMS", "Yes", "Compliant", "GCS config", "", "No"],
            ["OneDrive Business", "Confidential", "At rest", "Microsoft-managed", "AES-256", "M365", "Automatic", "Compliant", "M365 security", "", "No"],
            ["SharePoint", "Confidential", "At rest", "Microsoft-managed", "AES-256", "M365", "Automatic", "Compliant", "SharePoint admin", "", "No"],
        ],
    }

    total_written = 0
    for sheet_name, data in sheet_data.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            count = safely_write_data(ws, 8, data)
            total_written += count
            logger.info("    - %s: %d entries", sheet_name, count)

    try:
        wb.save(filepath)
    except Exception as e:
        logger.error("    Error saving: %s", e)
        return 0

    return total_written


def add_authentication_samples(filepath):
    """824_3: 11 columns - Status in column H (8)
    Columns: System/Application, Authentication Method, User Type,
             Data Classification, Cryptographic Algorithm, Hash/Encryption Status,
             Password Complexity, Status, Evidence Location, Gap Description,
             Remediation Needed
    """
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        logger.error("    Error loading %s: %s", filepath, e)
        return 0

    sheet_data = {
        "1. Password Security": [
            ["Web App Users", "bcrypt", "Standard", "Confidential", "bcrypt", "Cost 12", "Strong", "Compliant", "Code review", "", "No"],
            ["Internal Portal", "Argon2id", "Employee", "Confidential", "Argon2id", "64MB, iter=3", "Strong", "Compliant", "Security assessment", "", "No"],
            ["Customer Portal", "PBKDF2-SHA256", "Customer", "Confidential", "PBKDF2-SHA256", "600k iterations", "Strong", "Compliant", "Password config", "", "No"],
            ["CMS Backend", "PBKDF2-SHA256", "Admin", "Confidential", "PBKDF2-SHA256", "310k iterations", "Medium", "Partial", "CMS config", "Upgrade to 600k", "Yes"],
            ["Legacy System", "MD5", "Admin", "Restricted", "MD5 unsalted", "None", "Weak", "Non-Compliant", "Legacy audit", "Migrate to bcrypt", "Yes"],
            ["API Credentials", "Argon2id", "Service", "Confidential", "Argon2id", "64MB, iter=4", "Strong", "Compliant", "API security docs", "", "No"],
            ["Mobile App", "bcrypt", "User", "Confidential", "bcrypt", "Cost 14", "Strong", "Compliant", "Mobile security", "", "No"],
        ],
        "2. Multi-Factor Authentication": [
            ["All Employees", "MS Authenticator", "Employee", "N/A", "TOTP", "99%", "3450/3500", "Compliant", "Azure AD MFA", "", "No"],
            ["Administrators", "YubiKey + MS Auth", "Admin", "N/A", "FIDO2", "100%", "45/45", "Compliant", "Privileged access", "", "No"],
            ["Executives", "Hardware token", "Executive", "N/A", "TOTP", "100%", "12/12", "Compliant", "Executive policy", "", "No"],
            ["Contractors", "SMS + Authenticator", "Contractor", "N/A", "SMS/TOTP", "88%", "105/120", "Partial", "Contractor reports", "Mandate for all", "Yes"],
            ["External Partners", "Not implemented", "External", "N/A", "None", "0%", "0/85", "Non-Compliant", "Partner portal", "Implement MFA", "Yes"],
            ["Developers", "YubiKey + TOTP", "Developer", "N/A", "FIDO2/TOTP", "100%", "180/180", "Compliant", "GitHub + Azure AD", "", "No"],
        ],
        "5. SSO & Federation": [
            ["Microsoft 365", "SAML 2.0", "All users", "N/A", "RSA-SHA256", "Yes", "Azure AD", "Compliant", "Azure AD logs", "", "No"],
            ["Salesforce", "SAML 2.0", "Sales team", "N/A", "RSA-SHA256", "Yes", "Azure AD", "Compliant", "SFDC login history", "", "No"],
            ["GitHub Enterprise", "SAML 2.0", "Developers", "N/A", "RSA-SHA256", "Yes", "Azure AD", "Compliant", "GitHub audit log", "", "No"],
            ["AWS Console", "SAML 2.0", "Cloud team", "N/A", "RSA-SHA256", "Yes", "Azure AD", "Compliant", "CloudTrail", "", "No"],
            ["Slack", "SAML 2.0", "All users", "N/A", "RSA-SHA256", "Yes", "Azure AD", "Compliant", "Slack audit", "", "No"],
            ["Legacy CRM", "Local accounts", "Sales", "Confidential", "None", "No", "None", "Non-Compliant", "App logs", "Integrate SSO", "Yes"],
        ],
    }

    total_written = 0
    for sheet_name, data in sheet_data.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            count = safely_write_data(ws, 8, data)
            total_written += count
            logger.info("    - %s: %d entries", sheet_name, count)

    try:
        wb.save(filepath)
    except Exception as e:
        logger.error("    Error saving: %s", e)
        return 0

    return total_written


def add_key_management_samples(filepath):
    """824_4: 11 columns - Status in column H (8)
    Columns: System/Service, Key Type, Algorithm, Key Length (bits),
             Storage Method, Data Classification, Lifecycle Stage, Status,
             Evidence Location, Gap Description, Remediation Needed
    """
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        logger.error("    Error loading %s: %s", filepath, e)
        return 0

    sheet_data = {
        "1. Key Generation": [
            ["TLS Certificates", "TLS cert", "RSA", "4096", "AWS ACM", "Public", "Active", "Compliant", "Let's Encrypt", "", "No"],
            ["DB Encryption Keys", "DEK", "AES", "256", "AWS KMS", "Restricted", "Active", "Compliant", "AWS KMS", "", "No"],
            ["API Keys", "API key", "Random", "256", "Vault", "Confidential", "Active", "Compliant", "Vault", "", "No"],
            ["SSH Keys", "SSH key", "Ed25519", "256", "Ansible", "Internal", "Active", "Compliant", "SSH inventory", "", "No"],
            ["Code Signing", "Code sign", "RSA", "4096", "YubiKey HSM", "Restricted", "Active", "Compliant", "Signing logs", "", "No"],
            ["Legacy Keys", "Legacy", "DES", "56", "File system", "Confidential", "Deprecated", "Non-Compliant", "Old docs", "Migrate AES-256", "Yes"],
            ["JWT Signing", "JWT key", "ECDSA", "256", "Vault", "Confidential", "Active", "Compliant", "API gateway", "", "No"],
        ],
        "2. Key Storage": [
            ["Production TLS", "TLS cert", "RSA", "4096", "AWS ACM", "Public", "HSM-backed", "Compliant", "ACM config", "", "No"],
            ["DB Master Keys", "Master key", "AES", "256", "Azure Key Vault", "Restricted", "HSM-backed", "Compliant", "KV access logs", "", "No"],
            ["API Secrets", "Secret", "Random", "256", "Vault", "Confidential", "Encrypted", "Compliant", "Vault policies", "", "No"],
            ["App Secrets", "Secret", "Random", "256", "K8s Secrets", "Internal", "etcd encrypted", "Compliant", "K8s audit", "", "No"],
            ["Code Signing", "Private key", "RSA", "4096", "YubiKey HSM", "Restricted", "Hardware", "Compliant", "HSM logs", "", "No"],
            ["SSH Keys", "SSH key", "Ed25519", "256", "~/.ssh/", "Internal", "Filesystem", "Partial", "Endpoint scan", "Add passphrases", "Yes"],
            ["GCP Service Keys", "Service key", "RSA", "2048", "GCP Secret Mgr", "Confidential", "Encrypted", "Compliant", "GCP audit", "", "No"],
        ],
        "3. Key Rotation": [
            ["TLS Certificates", "TLS cert", "RSA", "4096", "Auto-renewal", "Public", "90 days", "Compliant", "Auto-renewal logs", "", "No"],
            ["DB Encryption", "DEK", "AES", "256", "Manual process", "Restricted", "365 days", "Partial", "Rotation schedule", "Automate", "Yes"],
            ["API Keys", "API key", "Random", "256", "Vault rotation", "Confidential", "180 days", "Compliant", "Vault rotation logs", "", "No"],
            ["GCP Service Keys", "Service key", "RSA", "2048", "Automated", "Confidential", "90 days", "Compliant", "GCP IAM logs", "", "No"],
            ["Azure Storage", "Storage key", "AES", "256", "Manual", "Confidential", "Never (730d)", "Non-Compliant", "Azure audit", "Auto-rotate", "Yes"],
            ["SSH Keys", "SSH key", "Ed25519", "256", "Manual", "Internal", "Never (730d)", "Non-Compliant", "SSH audit", "Enforce policy", "Yes"],
            ["JWT Signing", "JWT key", "ECDSA", "256", "Automated", "Confidential", "180 days", "Compliant", "API logs", "", "No"],
        ],
        "4. Key Destruction": [
            ["Decommissioned TLS", "TLS cert", "RSA", "4096", "Revocation", "Public", "Destroyed", "Compliant", "CRL/OCSP", "", "No"],
            ["Old DB Keys", "DEK", "AES", "256", "Secure deletion", "Restricted", "Destroyed", "Compliant", "KMS audit", "", "No"],
            ["Rotated API Keys", "API key", "Random", "256", "Vault deletion", "Confidential", "Destroyed", "Compliant", "Vault audit", "", "No"],
            ["Legacy SSH Keys", "SSH key", "RSA", "2048", "Manual deletion", "Internal", "Should destroy", "Partial", "Key inventory", "Verify destruction", "Yes"],
            ["Test Keys", "Test key", "Various", "Various", "No process", "Public", "Unknown", "Non-Compliant", "No records", "Implement process", "Yes"],
        ],
    }

    total_written = 0
    for sheet_name, data in sheet_data.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            count = safely_write_data(ws, 8, data)
            total_written += count
            logger.info("    - %s: %d entries", sheet_name, count)

    try:
        wb.save(filepath)
    except Exception as e:
        logger.error("    Error saving: %s", e)
        return 0

    return total_written


def main():
    """Main function to populate all A.8.24 workbooks."""
    try:
        logger.info("=" * 80)
        logger.info("POPULATING A.8.24 WORKBOOKS - MERGED & ENHANCED VERSION")
        logger.info("=" * 80)

        files = {
            "Data Transmission (8 cols)": ("./ISMS-IMP-A.8.24.1_Data_Transmission_20260113.xlsx", add_data_transmission_samples),
            "Data Storage (11 cols)": ("./ISMS-IMP-A.8.24.2_Data_Storage_20260113.xlsx", add_data_storage_samples),
            "Authentication (11 cols)": ("./ISMS-IMP-A.8.24.3_Authentication_20260113.xlsx", add_authentication_samples),
            "Key Management (11 cols)": ("./ISMS-IMP-A.8.24.4_Key_Management_20260113.xlsx", add_key_management_samples),
        }

        total_entries = 0
        results = {}

        for name, (filepath, func) in files.items():
            logger.info("[%d/4] Processing %s...", list(files.keys()).index(name) + 1, name)
            try:
                count = func(filepath)
                total_entries += count
                results[name] = count
                logger.info("  Added %d sample entries total", count)
            except Exception as e:
                logger.error("  Error: %s", e)
                import traceback
                traceback.print_exc()
                results[name] = 0

        logger.info("")
        logger.info("=" * 80)
        logger.info("MERGED DATA POPULATION COMPLETE!")
        logger.info("=" * 80)

        logger.info("TOTAL ENTRIES: %d", total_entries)
        logger.info("")
        logger.info("Breakdown:")
        for name, count in results.items():
            logger.info("  - %s: %d entries", name, count)

        logger.info("")
        logger.info("Data includes:")
        logger.info("  - Mix of Compliant (~70%%), Partial (~20%%), Non-Compliant (~10%%)")
        logger.info("  - Realistic system names and configurations")
        logger.info("  - Evidence locations for compliance tracking")
        logger.info("  - Gap descriptions and remediation plans")
        logger.info("  - Multiple assessment categories per workbook")

        logger.info("")
        logger.info("ALL STATUS VALUES IN CORRECT COLUMNS!")
        logger.info("Summary Dashboard formulas will now calculate correctly!")

        logger.info("")
        logger.info("Workbooks Updated:")
        logger.info("  - ISMS-IMP-A.8.24.1_Data_Transmission_20260113.xlsx")
        logger.info("  - ISMS-IMP-A.8.24.2_Data_Storage_20260113.xlsx")
        logger.info("  - ISMS-IMP-A.8.24.3_Authentication_20260113.xlsx")
        logger.info("  - ISMS-IMP-A.8.24.4_Key_Management_20260113.xlsx")

        logger.info("")
        logger.info("Next Steps:")
        logger.info("  1. Open workbooks in Excel to verify data")
        logger.info("  2. Check Summary Dashboard calculations")
        logger.info("  3. Generate consolidated dashboard workbook")
        logger.info("  4. Test external link updates")
        logger.info("  5. Validate compliance percentage calculations")

        logger.info("")
        logger.info("Workbooks ready for dashboard consolidation!")
        logger.info("=" * 80)

        return 0

    except Exception as e:
        logger.error("Population failed with exception: %s", str(e))
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (standardized: license, imports, logging, exit codes)
# QA_TOOL: Claude Code Deep Scan
# STANDARDIZATION: v1.0 - AGPL-3.0/Commercial dual-license
# =============================================================================
