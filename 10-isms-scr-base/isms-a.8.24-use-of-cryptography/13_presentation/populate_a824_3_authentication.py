#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
ISMS-IMP-A.8.24.3 - Authentication Assessment
Comprehensive Data Population for CISO Presentation

Generated: 2026-01-13
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


def safely_write_data(ws, start_row, data):
    """Safely write data to worksheet, handling merged cells."""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except Exception as e:  # Skip merged/protected cells
                continue
        entries_written += 1
    return entries_written


def populate_assessment_sheets(wb):
    """Populate all assessment sheets."""
    sheet_data = {
        "1. Password Security": [
            ["Active Directory", "NIST SP 800-63B compliant", "Argon2id", "12 characters minimum", "Yes - 90 days", "GPO enforced", "Password policy GPO", "Compliant", "AD audit logs", "", "No"],
            ["Azure AD", "NIST compliant + MFA", "Azure AD hash", "12 characters + complexity", "Yes - Conditional", "Conditional Access", "Azure AD security", "Compliant", "Azure AD logs", "", "No"],
            ["Linux PAM", "NIST compliant", "SHA-512 + salt", "14 characters", "Yes - 90 days", "PAM config", "PAM config files", "Compliant", "/etc/pam.d/ audit", "", "No"],
            ["Database Accounts", "Application-specific", "SCRAM-SHA-256", "16 characters", "Yes - 180 days", "DB password policy", "DB security audit", "Compliant", "DB password logs", "", "No"],
            ["Application Users", "OWASP compliant", "bcrypt (cost 12)", "12 characters + 2FA", "Yes - 120 days", "App password policy", "App security config", "Compliant", "App audit logs", "", "No"],
            ["Legacy Applications", "Basic requirements", "MD5 hash", "8 characters", "No expiry", "No policy", "Legacy app list", "Non-Compliant", "Legacy audit", "Migrate or decommission", "Yes"],
            ["Service Accounts (AD)", "Strong passwords", "Kerberos AES-256", "20+ random chars", "No expiry + monitoring", "Managed Service Accounts", "AD service account policy", "Compliant", "AD service logs", "", "No"],
            ["API Keys", "Cryptographically strong", "HMAC-SHA-256", "32+ bytes entropy", "Yes - 180 days", "Vault rotation", "API key management", "Compliant", "Vault audit logs", "", "No"],
        ],

        "2. Multi-Factor Authentication": [
            ["Azure AD MFA", "TOTP + Push + FIDO2", "100% admin, 95% users", "Authenticator app", "Yes - Conditional Access", "Azure AD MFA", "Azure MFA report", "Compliant", "Azure AD MFA logs", "", "No"],
            ["VPN Access MFA", "TOTP + Push", "100% required", "Duo Security", "Yes - Always", "Duo MFA", "Duo admin dashboard", "Compliant", "Duo authentication logs", "", "No"],
            ["SSH Bastion MFA", "FIDO2 hardware keys", "100% required", "YubiKey 5", "Yes - Always", "PAM + Yubikey", "SSH bastion config", "Compliant", "SSH audit logs", "", "No"],
            ["Admin Portal MFA", "TOTP + Backup codes", "100% required", "Google Authenticator", "Yes - Always", "App-level MFA", "Admin portal config", "Compliant", "Admin access logs", "", "No"],
            ["AWS Console MFA", "Virtual MFA + U2F", "100% IAM users", "AWS MFA", "Yes - Always", "IAM MFA policies", "IAM credential report", "Compliant", "CloudTrail MFA events", "", "No"],
            ["Critical Systems", "FIDO2 + Biometric", "100% required", "Windows Hello + YubiKey", "Yes - Always", "Conditional Access", "Critical system policy", "Compliant", "Access audit logs", "", "No"],
            ["Developer Systems", "TOTP", "85% adoption", "Various", "Recommended", "Not enforced", "Dev MFA report", "Partial", "Dev access logs", "Enforce 100% MFA", "Yes"],
            ["Legacy Systems", "No MFA", "0% coverage", "N/A", "Not supported", "N/A", "Legacy system list", "Non-Compliant", "Legacy access logs", "Implement MFA proxy", "Yes"],
        ],

        "3. Certificate-Based Auth": [
            ["Smart Card Logon (PIV)", "X.509 certificates", "100% executives", "PIV smart cards", "Annual cert renewal", "AD Certificate Services", "PIV deployment docs", "Compliant", "Smart card logs", "", "No"],
            ["Client Certificates (mTLS)", "X.509 certificates", "100% API clients", "Client certs", "Certificate rotation", "Internal PKI", "mTLS architecture", "Compliant", "PKI audit logs", "", "No"],
            ["Machine Certificates", "X.509 certificates", "100% managed devices", "MDM-issued", "Automatic renewal", "MDM PKI integration", "MDM certificate policy", "Compliant", "MDM cert logs", "", "No"],
            ["VPN Certificates", "X.509 + username", "Partner VPN only", "Client certificates", "Annual renewal", "VPN CA", "VPN cert config", "Compliant", "VPN authentication logs", "", "No"],
            ["Email Signing (S/MIME)", "X.509 certificates", "85% executives", "S/MIME certs", "Annual renewal", "Public CA + Internal", "S/MIME deployment", "Partial", "Certificate store audit", "Increase to 100%", "Yes"],
            ["Code Signing", "X.509 + HSM", "100% releases", "EV code signing", "Annual renewal", "YubiKey HSM", "Code signing procedure", "Compliant", "Signing logs", "", "No"],
        ],

        "4. Service Accounts": [
            ["Azure Managed Identities", "Azure AD workload identity", "100% Azure resources", "Managed Identity", "N/A - Automatic", "Azure RBAC", "Azure MI configuration", "Compliant", "Azure activity logs", "", "No"],
            ["AWS IAM Roles", "IAM role assumption", "100% AWS resources", "IAM Roles", "N/A - Temporary creds", "IAM policies", "IAM role documentation", "Compliant", "CloudTrail AssumeRole", "", "No"],
            ["GCP Service Accounts", "GCP service account", "100% GCP resources", "Service Account keys", "Key rotation 90 days", "GCP IAM", "GCP SA inventory", "Compliant", "Cloud Audit Logs", "", "No"],
            ["AD Group Managed Service Accounts", "gMSA", "80% Windows services", "gMSA", "N/A - AD managed", "AD gMSA", "gMSA deployment", "Partial", "AD service logs", "Migrate remaining services", "Yes"],
            ["Kubernetes Service Accounts", "K8s SA + RBAC", "100% pods", "JWT tokens", "Automatic rotation", "K8s RBAC", "K8s SA policies", "Compliant", "K8s audit logs", "", "No"],
            ["Database Service Accounts", "DB-specific", "100% apps", "Strong passwords", "180 days", "Secret rotation", "DB account inventory", "Partial", "DB audit logs", "Implement secret manager", "Yes"],
            ["Legacy Service Accounts", "Hardcoded passwords", "Legacy apps", "Static passwords", "No rotation", "No management", "Legacy SA list", "Non-Compliant", "Legacy app audit", "Modernize authentication", "Yes"],
        ],

        "5. SSO & Federation": [
            ["Azure AD SSO", "SAML 2.0 + OpenID Connect", "95% SaaS apps", "Azure AD", "Conditional Access enforced", "Azure AD SSO", "Azure AD app gallery", "Compliant", "Azure AD sign-ins", "", "No"],
            ["Okta SSO", "SAML 2.0 + OAuth 2.0", "100% federated apps", "Okta", "MFA + Adaptive Auth", "Okta Universal Directory", "Okta app integrations", "Compliant", "Okta system logs", "", "No"],
            ["ADFS Federation", "WS-Federation + SAML", "100% AD-integrated", "ADFS", "MFA enforced", "ADFS + Azure AD", "ADFS architecture", "Compliant", "ADFS audit logs", "", "No"],
            ["Google Workspace SSO", "SAML 2.0", "100% Google apps", "Google SSO", "Security key enforced", "Google Admin", "Google SSO config", "Compliant", "Google Admin logs", "", "No"],
            ["AWS SSO (IAM Identity Center)", "SAML 2.0", "100% AWS accounts", "AWS SSO", "MFA required", "AWS Organizations", "AWS SSO configuration", "Compliant", "CloudTrail SSO events", "", "No"],
            ["Custom Apps with SSO", "SAML 2.0 + OIDC", "80% custom apps", "Various IdPs", "MFA where supported", "App-specific", "Custom app inventory", "Partial", "App authentication logs", "Complete SSO integration", "Yes"],
            ["Legacy Apps (No SSO)", "Local authentication", "20% legacy apps", "N/A", "N/A", "Local credentials", "Legacy app list", "Non-Compliant", "Local auth logs", "Implement SSO proxy", "Yes"],
        ],
    }

    total_written = 0
    for sheet_name, data in sheet_data.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            count = safely_write_data(ws, 8, data)
            total_written += count
            logger.info("    %s: %d entries", sheet_name, count)
    return total_written


def populate_evidence_register(wb):
    """Populate Evidence Register."""
    if "Evidence Register" not in wb.sheetnames:
        return 0
    ws = wb["Evidence Register"]
    evidence_data = [
        ["EVD-824.3-001", "Active Directory Password Policy GPO", "GPO Export", "1. Password Security", "AD/GPO/Password", "2026-01-10", "Windows Admin", "3 years", "Current"],
        ["EVD-824.3-002", "Azure AD Password Policy Configuration", "Screenshot", "1. Password Security", "Azure/AAD/Security", "2026-01-10", "Cloud Admin", "2 years", "Current"],
        ["EVD-824.3-003", "Linux PAM Configuration Audit", "Config Files", "1. Password Security", "Linux/PAM/Config", "2026-01-09", "Linux Admin", "2 years", "Current"],
        ["EVD-824.3-004", "Password Security Standard v3.2", "Standard Document", "1. Password Security", "SharePoint/Standards", "2025-08-01", "Security Team", "5 years", "Current"],
        ["EVD-824.3-005", "Azure MFA Adoption Report", "Compliance Report", "2. Multi-Factor Authentication", "Azure/MFA/Reports", "2026-01-10", "Security Team", "1 year", "Current"],
        ["EVD-824.3-006", "Duo Security Deployment Configuration", "Config Export", "2. Multi-Factor Authentication", "Duo/Admin/Config", "2026-01-09", "Security Team", "2 years", "Current"],
        ["EVD-824.3-007", "YubiKey Deployment Inventory", "Inventory List", "2. Multi-Factor Authentication", "SharePoint/Inventory/MFA", "2026-01-08", "IT Admin", "2 years", "Current"],
        ["EVD-824.3-008", "MFA Policy Document v2.5", "Policy", "2. Multi-Factor Authentication", "SharePoint/Policies/MFA", "2025-09-01", "Security Team", "5 years", "Current"],
        ["EVD-824.3-009", "PIV Smart Card Deployment Guide", "Technical Guide", "3. Certificate-Based Auth", "SharePoint/Guides/PIV", "2025-10-15", "Security Team", "3 years", "Current"],
        ["EVD-824.3-010", "PKI Certificate Policy", "Policy Document", "3. Certificate-Based Auth", "SharePoint/Policies/PKI", "2025-06-01", "Security Team", "5 years", "Current"],
        ["EVD-824.3-011", "mTLS Architecture Documentation", "Architecture Doc", "3. Certificate-Based Auth", "Confluence/mTLS", "2025-11-01", "Platform Team", "3 years", "Current"],
        ["EVD-824.3-012", "S/MIME Certificate Deployment Report", "Deployment Report", "3. Certificate-Based Auth", "SharePoint/Certificates", "2026-01-09", "Email Admin", "2 years", "Current"],
        ["EVD-824.3-013", "Azure Managed Identity Implementation", "Technical Doc", "4. Service Accounts", "Azure/Documentation", "2025-12-01", "Cloud Team", "3 years", "Current"],
        ["EVD-824.3-014", "AWS IAM Role Best Practices", "Best Practices Doc", "4. Service Accounts", "AWS/Documentation", "2025-11-15", "Cloud Team", "3 years", "Current"],
        ["EVD-824.3-015", "gMSA Deployment Status Report", "Status Report", "4. Service Accounts", "AD/gMSA/Reports", "2026-01-08", "Windows Admin", "2 years", "Current"],
        ["EVD-824.3-016", "Service Account Inventory Q1 2026", "Inventory", "4. Service Accounts", "SharePoint/Inventory/SA", "2026-01-10", "Security Team", "2 years", "Current"],
        ["EVD-824.3-017", "Azure AD SSO Configuration", "Config Export", "5. SSO & Federation", "Azure/AAD/SSO", "2026-01-10", "Cloud Admin", "2 years", "Current"],
        ["EVD-824.3-018", "Okta Application Integration List", "Integration List", "5. SSO & Federation", "Okta/Admin/Apps", "2026-01-09", "IT Admin", "2 years", "Current"],
        ["EVD-824.3-019", "ADFS Federation Configuration", "Config Documentation", "5. SSO & Federation", "ADFS/Configuration", "2025-12-15", "Identity Team", "3 years", "Current"],
        ["EVD-824.3-020", "SSO Implementation Standard", "Standard Document", "5. SSO & Federation", "SharePoint/Standards/SSO", "2025-09-01", "Security Team", "5 years", "Current"],
    ]
    count = safely_write_data(ws, 10, evidence_data)
    logger.info("    Evidence Register: %d evidence documents", count)
    return count


def populate_approval_signoff(wb):
    """Populate Approval Sign-Off."""

    if "Approval Sign-Off" not in wb.sheetnames:
        return 0

    ws = wb["Approval Sign-Off"]

    today = datetime.now()

    # Helper function to safely write to cells (handles merged cells)
    def safe_write(cell_ref, value):
        try:
            cell = ws[cell_ref]
            if not isinstance(cell, MergedCell):
                cell.value = value
        except Exception as e:  # Skip merged/protected cells
            pass

    # Assessment completion
    safe_write("B5", "Alex Thompson")
    safe_write("B6", "Identity & Access Management Engineer")
    safe_write("B7", "Information Security")
    safe_write("B8", "alex.thompson@company.com")
    safe_write("B9", today.strftime("%Y-%m-%d"))

    # Technical Review
    safe_write("B14", "Lisa Martinez")
    safe_write("B15", (today + timedelta(days=2)).strftime("%Y-%m-%d"))
    safe_write("B16", "Reviewed and verified all authentication controls. Legacy system SSO integration needs attention.")

    # Security Review
    safe_write("B21", "Michael Chen")
    safe_write("B22", (today + timedelta(days=3)).strftime("%Y-%m-%d"))
    safe_write("B23", "MFA coverage excellent. Focus on completing developer MFA enforcement and legacy app modernization.")

    # Management Approval
    safe_write("B28", "Jennifer Williams")
    safe_write("B29", (today + timedelta(days=5)).strftime("%Y-%m-%d"))
    safe_write("B30", "Approved")
    safe_write("B31", "Approved. Strong authentication posture overall.")

    # Next Review
    safe_write("B36", (today + timedelta(days=90)).strftime("%Y-%m-%d"))
    safe_write("B37", "Lisa Martinez")
    safe_write("B38", "Q2 2026 follow-up on legacy application authentication modernization")

    logger.info("    Approval Sign-Off: Complete")
    return 1


def main():
    """Main function to populate the Authentication assessment workbook."""
    try:
        if len(sys.argv) < 2:
            logger.error("Usage: python3 populate_a824_3_authentication.py <workbook.xlsx>")
            return 1

        filepath = sys.argv[1]

        logger.info("=" * 80)
        logger.info("ISMS-IMP-A.8.24.3 - Authentication Assessment")
        logger.info("Comprehensive Data Population for CISO Presentation")
        logger.info("=" * 80)

        try:
            wb = load_workbook(filepath)
            logger.info("Loading: %s", filepath)
            logger.info("Sheets found: %d", len(wb.sheetnames))

            logger.info("[1/3] Populating 5 Assessment Sheets...")
            assessment_count = populate_assessment_sheets(wb)

            logger.info("[2/3] Populating Evidence Register...")
            evidence_count = populate_evidence_register(wb)

            logger.info("[3/3] Populating Approval Sign-Off...")
            approval_count = populate_approval_signoff(wb)

            wb.save(filepath)
            logger.info("Saved: %s", filepath)

            logger.info("=" * 80)
            logger.info("DATA POPULATION COMPLETE")
            logger.info("=" * 80)
            logger.info("Summary:")
            logger.info("  - Assessment Entries: %d", assessment_count)
            logger.info("  - Evidence Documents: %d", evidence_count)
            logger.info("  - Total Data Points: %d", assessment_count + evidence_count)
            logger.info("CISO-Ready: Professional authentication assessment")
            logger.info("=" * 80)

            return 0

        except Exception as e:
            logger.error("Error: %s", e)
            import traceback
            traceback.print_exc()
            return 1

    except Exception as e:
        logger.error("Population failed with exception: %s", str(e))
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (standardized: license, imports, logging, exit codes)
# QA_TOOL: Claude Code Deep Scan
# STANDARDIZATION: v1.0 - AGPL-3.0/Commercial dual-license
# =============================================================================
