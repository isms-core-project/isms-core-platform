#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
ISMS-IMP-A.8.24.2 - Data Storage Assessment
Comprehensive Data Population Script for CISO Presentation

Populates ALL sheets with production-quality demo data:
- 7 Assessment sheets (Mobile, Laptops, Servers, Databases, Cloud, Backups, Media)
- Summary Dashboard (auto-calculated)
- Evidence Register (25+ evidence documents)
- Approval Sign-Off (complete workflow)

Usage:
    python3 populate_a824_2_data_storage.py ISMS-IMP-A.8.24.2.xlsx

Generated: 2026-01-13
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


def safely_write_data(ws, start_row, data):
    """Safely write data to worksheet, handling merged cells."""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except Exception as e:  # Skip merged/protected cells
                continue
        entries_written += 1
    return entries_written


def populate_assessment_sheets(wb):
    """Populate all 7 assessment sheets."""

    # Columns: Device/System, Data Classification, Encryption Status, Encryption Type,
    #          Algorithm & Key Size, Key Management, Key Rotation, Status, Evidence, Gap, Remediation

    sheet_data = {
        "1. Mobile Devices": [
            ["Corporate iPhones (500 units)", "Confidential", "iOS FileVault", "Hardware-backed", "AES-256", "MDM - MobileIron", "Yes", "Compliant", "MDM compliance report", "", "No"],
            ["Android Devices (300 units)", "Confidential", "Android FDE", "Hardware-backed", "AES-256", "MDM - Workspace ONE", "Yes", "Compliant", "Workspace ONE dashboard", "", "No"],
            ["Corporate iPads (200 units)", "Confidential", "iOS FileVault", "Hardware-backed", "AES-256", "MDM - Jamf Pro", "Yes", "Compliant", "Jamf compliance scan", "", "No"],
            ["Executive iPhones", "Restricted", "iOS + Container", "Hardware + App", "AES-256", "MDM + App wrapping", "Yes", "Compliant", "Executive device audit", "", "No"],
            ["BYOD iPhones (enrolled)", "Internal", "Containerization", "App-level", "AES-256", "MDM container only", "Yes", "Partial", "BYOD policy docs", "Increase enrollment", "Yes"],
            ["BYOD Android (enrolled)", "Internal", "Containerization", "App-level", "AES-256", "MDM container only", "Yes", "Partial", "BYOD enrollment report", "Training needed", "Yes"],
            ["Test Devices (50 units)", "Public", "Various/None", "Mixed/None", "Various", "No central management", "No", "Non-Compliant", "Test device inventory", "Enroll in MDM or decommission", "Yes"],
            ["Legacy Tablets", "Internal", "No encryption", "None", "None", "No management", "No", "Non-Compliant", "Legacy device list", "Replace or decommission", "Yes"],
        ],

        "2. Laptops & Workstations": [
            ["Windows 11 Laptops (800 units)", "Confidential", "BitLocker", "Full Disk", "AES-256-XTS", "AD + Intune recovery", "Annual", "Compliant", "BitLocker compliance dashboard", "", "No"],
            ["MacBook Pro (400 units)", "Confidential", "FileVault 2", "Full Disk", "AES-256-XTS", "MDM escrowed keys", "Annual", "Compliant", "Jamf encryption report", "", "No"],
            ["MacBook Air (250 units)", "Confidential", "FileVault 2", "Full Disk", "AES-256-XTS", "MDM escrowed keys", "Annual", "Compliant", "Jamf encryption report", "", "No"],
            ["Linux Workstations (100 units)", "Internal", "LUKS", "Full Disk", "AES-256-XTS", "Manual key backup", "Annual", "Partial", "Linux asset inventory", "Centralize key management", "Yes"],
            ["Executive Laptops", "Restricted", "BitLocker + TPM 2.0", "Full Disk + TPM", "AES-256-XTS", "Intune + Hardware TPM", "Quarterly", "Compliant", "Executive device report", "", "No"],
            ["Developer MacBooks", "Confidential", "FileVault 2", "Full Disk", "AES-256-XTS", "MDM managed", "Annual", "Compliant", "Dev team device audit", "", "No"],
            ["Windows 10 Desktops (500 units)", "Internal", "BitLocker", "Full Disk", "AES-256", "AD recovery", "Annual", "Compliant", "SCCM compliance", "", "No"],
            ["Contractor Laptops (80 units)", "Internal", "Various/None", "Mixed", "Various", "No enforcement", "No", "Non-Compliant", "Contractor device list", "Enforce encryption policy", "Yes"],
            ["Legacy Desktops (50 units)", "Internal", "No encryption", "None", "None", "None", "No", "Non-Compliant", "Legacy hardware list", "Upgrade or decommission", "Yes"],
        ],

        "3. Servers": [
            ["Production Linux (150 units)", "Restricted", "LUKS + dm-crypt", "Full Disk", "AES-256-XTS", "Ansible Vault", "Annual", "Compliant", "Server encryption audit", "", "No"],
            ["Production Windows (100 units)", "Restricted", "BitLocker", "Full Disk", "AES-256-XTS", "MBAM + AD", "Annual", "Compliant", "BitLocker server status", "", "No"],
            ["VMware VMs (300 units)", "Confidential", "VM Encryption", "VMDK-level", "AES-256-XTS", "vCenter + KMS", "Annual", "Compliant", "vSphere encryption report", "", "No"],
            ["Hyper-V VMs (150 units)", "Confidential", "Shielded VMs", "VM-level", "AES-256", "SCVMM + GuardedHost", "Annual", "Compliant", "Hyper-V audit", "", "No"],
            ["Docker Hosts (80 units)", "Internal", "Host LUKS", "Host disk", "AES-256", "Ansible managed", "Annual", "Compliant", "Docker host inventory", "", "No"],
            ["Kubernetes Nodes (120 units)", "Confidential", "etcd encryption", "etcd + disk", "AES-256-GCM", "K8s secrets + LUKS", "Quarterly", "Compliant", "K8s encryption config", "", "No"],
            ["Development Servers (50 units)", "Internal", "LUKS", "Full Disk", "AES-256", "Manual", "Annual", "Partial", "Dev server list", "Automate key rotation", "Yes"],
            ["Legacy Unix Servers (20 units)", "Internal", "No encryption", "None", "None", "None", "No", "Non-Compliant", "Legacy server inventory", "Migrate or decommission", "Yes"],
        ],

        "4. Databases": [
            ["PostgreSQL Production (10 nodes)", "Restricted", "TDE - pgcrypto", "Transparent", "AES-256-GCM", "AWS KMS", "Yes - Quarterly", "Compliant", "PostgreSQL TDE status", "", "No"],
            ["MySQL Customer DB (12 nodes)", "Confidential", "InnoDB TDE", "Transparent", "AES-256-CBC", "Azure Key Vault", "Yes - Quarterly", "Compliant", "MySQL encryption audit", "", "No"],
            ["SQL Server Finance (8 nodes)", "Restricted", "TDE + Always Encrypted", "Transparent + Column", "AES-256", "HSM-backed keys", "Yes - Quarterly", "Compliant", "SQL Server DMV reports", "", "No"],
            ["MongoDB Analytics (6 nodes)", "Internal", "Encrypted Storage Engine", "At-rest", "AES-256-GCM", "Self-managed keyfile", "No", "Partial", "MongoDB config audit", "Implement KMS rotation", "Yes"],
            ["Redis Cache Cluster", "Internal", "RDB/AOF Encryption", "Snapshot", "AES-256", "Key in config", "No", "Partial", "Redis configuration", "Move keys to KMS", "Yes"],
            ["Oracle Enterprise (4 nodes)", "Confidential", "TDE - Advanced Security", "Transparent", "AES-192", "Manual key management", "No", "Partial", "Oracle TDE audit", "Upgrade to AES-256 + HSM", "Yes"],
            ["Cassandra Cluster (8 nodes)", "Confidential", "At-rest Encryption", "SSTable", "AES-256", "JKS keystore", "Manual", "Partial", "Cassandra encryption.yaml", "Automate key rotation", "Yes"],
            ["MariaDB Prod (6 nodes)", "Confidential", "InnoDB TDE", "Transparent", "AES-256", "File-based keys", "Manual", "Partial", "MariaDB encryption status", "Implement KMS", "Yes"],
            ["DynamoDB Tables", "Confidential", "AWS managed", "Server-side", "AES-256", "AWS KMS", "Yes - Automatic", "Compliant", "DynamoDB encryption config", "", "No"],
            ["Test Databases", "Public", "None", "None", "None", "None", "No", "Non-Compliant", "Test DB inventory", "Implement encryption", "Yes"],
        ],

        "5. Cloud Storage": [
            ["AWS S3 Production", "Confidential", "SSE-KMS", "Server-side", "AES-256-GCM", "AWS KMS", "Yes - Automatic", "Compliant", "S3 bucket encryption audit", "", "No"],
            ["Azure Blob Storage", "Confidential", "SSE with Customer Keys", "Server-side", "AES-256", "Azure Key Vault", "Yes - Automatic", "Compliant", "Azure Storage encryption", "", "No"],
            ["GCP Cloud Storage", "Confidential", "CMEK", "Server-side", "AES-256-GCM", "Cloud KMS", "Yes - Automatic", "Compliant", "GCS encryption config", "", "No"],
            ["AWS EBS Volumes", "Restricted", "EBS Encryption", "Block-level", "AES-256-XTS", "AWS KMS", "Yes - Automatic", "Compliant", "EBS encryption status", "", "No"],
            ["Azure Managed Disks", "Restricted", "SSE + Customer Keys", "Disk-level", "AES-256-XTS", "Azure Key Vault", "Yes - Automatic", "Compliant", "Managed Disk encryption", "", "No"],
            ["GCP Persistent Disks", "Restricted", "CMEK", "Disk-level", "AES-256-XTS", "Cloud KMS", "Yes - Automatic", "Compliant", "Persistent Disk config", "", "No"],
            ["AWS EFS File Systems", "Confidential", "EFS Encryption", "File-level", "AES-256-GCM", "AWS KMS", "Yes - Automatic", "Compliant", "EFS encryption status", "", "No"],
            ["OneDrive for Business", "Confidential", "Microsoft-managed", "Server-side", "AES-256", "Microsoft managed", "Yes - Automatic", "Compliant", "OneDrive admin report", "", "No"],
            ["SharePoint Online", "Confidential", "Microsoft-managed", "Server-side", "AES-256", "Microsoft managed", "Yes - Automatic", "Compliant", "SharePoint encryption", "", "No"],
            ["Dropbox Business", "Internal", "Dropbox-managed", "Server-side", "AES-256", "Dropbox managed", "Yes - Automatic", "Compliant", "Dropbox admin console", "", "No"],
        ],

        "6. Backups": [
            ["Veeam Backup Production", "Restricted", "Encrypted backups", "Full backup", "AES-256", "Veeam password + KMS", "Annual", "Compliant", "Veeam backup reports", "", "No"],
            ["Commvault Enterprise", "Restricted", "Encrypted backups", "Full backup", "AES-256", "Commvault key manager", "Annual", "Compliant", "Commvault encryption audit", "", "No"],
            ["AWS Backup", "Confidential", "AWS managed encryption", "Backup vaults", "AES-256-GCM", "AWS KMS", "Yes - Automatic", "Compliant", "AWS Backup config", "", "No"],
            ["Azure Backup", "Confidential", "Azure-managed", "Recovery vaults", "AES-256", "Azure Key Vault", "Yes - Automatic", "Compliant", "Azure Backup reports", "", "No"],
            ["Database Backup (SQL)", "Restricted", "TDE backup encryption", "Native backup", "AES-256", "SQL certificate", "Annual", "Compliant", "SQL backup audit", "", "No"],
            ["File Server Backup", "Confidential", "Encrypted tapes", "Tape backup", "AES-256", "Tape library encryption", "Annual", "Compliant", "Tape library logs", "", "No"],
            ["Offsite Backup (Iron Mountain)", "Restricted", "Encrypted media", "Physical tapes", "AES-256", "Key escrow", "Annual", "Compliant", "Iron Mountain inventory", "", "No"],
            ["Personal Backup (Time Machine)", "Internal", "FileVault encrypted", "Disk backup", "AES-256", "User-managed", "No", "Partial", "Backup policy", "Enforce corporate backup", "Yes"],
            ["Legacy Backup System", "Internal", "No encryption", "None", "None", "None", "No", "Non-Compliant", "Legacy backup inventory", "Migrate to encrypted solution", "Yes"],
        ],

        "7. Removable Media": [
            ["Encrypted USB Drives (IronKey)", "Confidential", "Hardware encrypted", "Full device", "AES-256", "Hardware-based", "N/A", "Compliant", "IronKey deployment list", "", "No"],
            ["Encrypted USB Drives (Kingston)", "Confidential", "Hardware encrypted", "Full device", "AES-256", "Hardware keypad", "N/A", "Compliant", "Kingston inventory", "", "No"],
            ["BitLocker To Go USB", "Internal", "BitLocker To Go", "Full device", "AES-256", "AD recovery", "N/A", "Compliant", "BitLocker To Go policy", "", "No"],
            ["External HDDs (encrypted)", "Confidential", "Hardware encrypted", "Full disk", "AES-256", "Hardware PIN", "N/A", "Compliant", "External HDD inventory", "", "No"],
            ["Corporate SD Cards", "Internal", "VeraCrypt", "Container", "AES-256", "User password", "N/A", "Partial", "SD card policy", "Enforce hardware encryption", "Yes"],
            ["DVD/CD Media (archived)", "Internal", "Password-protected", "File-level", "Various", "Document password", "N/A", "Partial", "Optical media inventory", "Phase out optical media", "Yes"],
            ["Standard USB Drives", "Public", "No encryption", "None", "None", "None", "N/A", "Non-Compliant", "USB policy violation logs", "Block standard USB via GPO", "Yes"],
            ["Personal USB Drives", "N/A", "Blocked by policy", "N/A", "N/A", "N/A", "N/A", "Compliant", "DLP USB blocking rules", "", "No"],
        ],
    }

    total_written = 0
    for sheet_name, data in sheet_data.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            count = safely_write_data(ws, 8, data)
            total_written += count
            logger.info("    %s: %d entries", sheet_name, count)

    return total_written


def populate_evidence_register(wb):
    """Populate Evidence Register."""

    if "Evidence Register" not in wb.sheetnames:
        return 0

    ws = wb["Evidence Register"]

    evidence_data = [
        ["EVD-824.2-001", "MDM Compliance Report - Mobile Devices", "Compliance Report", "1. Mobile Devices", "MobileIron/Reports", "2026-01-10", "IT Admin", "2 years", "Current"],
        ["EVD-824.2-002", "iOS Encryption Policy Configuration", "Configuration Doc", "1. Mobile Devices", "SharePoint/MDM/iOS", "2026-01-10", "Security Team", "3 years", "Current"],
        ["EVD-824.2-003", "Android Encryption Compliance Scan", "Scan Report", "1. Mobile Devices", "Workspace ONE/Scans", "2026-01-09", "IT Admin", "1 year", "Current"],
        ["EVD-824.2-004", "BYOD Policy Document v2.1", "Policy", "1. Mobile Devices", "SharePoint/Policies/BYOD", "2025-06-01", "HR & Security", "5 years", "Current"],
        ["EVD-824.2-005", "BitLocker Deployment Status Report", "Deployment Report", "2. Laptops & Workstations", "Intune/Reports", "2026-01-10", "Windows Admin", "2 years", "Current"],
        ["EVD-824.2-006", "FileVault 2 Compliance - Mac Fleet", "Compliance Report", "2. Laptops & Workstations", "Jamf/Encryption", "2026-01-10", "Mac Admin", "2 years", "Current"],
        ["EVD-824.2-007", "BitLocker Recovery Key Escrow Audit", "Audit Report", "2. Laptops & Workstations", "AD/BitLocker-Keys", "2025-12-15", "Security Team", "3 years", "Current"],
        ["EVD-824.2-008", "Linux LUKS Encryption Standards", "Technical Standard", "2. Laptops & Workstations", "SharePoint/Standards", "2025-08-01", "Linux Team", "5 years", "Current"],
        ["EVD-824.2-009", "Server Encryption Audit Q4 2025", "Audit Report", "3. Servers", "SharePoint/Audits/Servers", "2025-12-20", "Internal Audit", "7 years", "Current"],
        ["EVD-824.2-010", "VMware VM Encryption Configuration", "Config Export", "3. Servers", "vCenter/Configs", "2026-01-08", "VMware Admin", "2 years", "Current"],
        ["EVD-824.2-011", "Kubernetes etcd Encryption Config", "YAML Config", "3. Servers", "Git/k8s/encryption", "2025-11-15", "K8s Admin", "2 years", "Current"],
        ["EVD-824.2-012", "PostgreSQL TDE Implementation Guide", "Technical Doc", "4. Databases", "Confluence/Database/PG", "2025-09-20", "DBA Team", "3 years", "Current"],
        ["EVD-824.2-013", "MySQL TDE Status Report", "Status Report", "4. Databases", "SharePoint/DBA/MySQL", "2026-01-09", "DBA Team", "2 years", "Current"],
        ["EVD-824.2-014", "SQL Server TDE Configuration Audit", "Audit Report", "4. Databases", "SQL-Audit/TDE", "2025-12-10", "DBA Team", "3 years", "Current"],
        ["EVD-824.2-015", "AWS KMS Key Policy - Databases", "Policy Doc", "4. Databases", "AWS/KMS/Policies", "2025-10-15", "Cloud Team", "3 years", "Current"],
        ["EVD-824.2-016", "MongoDB Encryption at Rest Config", "Config File", "4. Databases", "Git/mongodb/encryption", "2025-11-01", "DBA Team", "2 years", "Current"],
        ["EVD-824.2-017", "S3 Bucket Encryption Compliance", "Compliance Report", "5. Cloud Storage", "AWS/S3/Encryption", "2026-01-10", "Cloud Team", "2 years", "Current"],
        ["EVD-824.2-018", "Azure Storage Encryption Config", "Config Screenshot", "5. Cloud Storage", "Azure/Storage/Security", "2026-01-10", "Cloud Team", "2 years", "Current"],
        ["EVD-824.2-019", "GCP CMEK Configuration Guide", "Technical Guide", "5. Cloud Storage", "GCP/Security/CMEK", "2025-10-01", "Cloud Team", "3 years", "Current"],
        ["EVD-824.2-020", "EBS Volume Encryption Status", "Status Report", "5. Cloud Storage", "AWS/EC2/EBS", "2026-01-09", "Cloud Team", "1 year", "Current"],
        ["EVD-824.2-021", "Veeam Backup Encryption Settings", "Config Export", "6. Backups", "Veeam/Config", "2026-01-08", "Backup Team", "2 years", "Current"],
        ["EVD-824.2-022", "AWS Backup Vault Encryption", "Screenshot", "6. Backups", "AWS/Backup/Config", "2026-01-10", "Cloud Team", "2 years", "Current"],
        ["EVD-824.2-023", "Backup Encryption Policy v3.0", "Policy", "6. Backups", "SharePoint/Policies/Backup", "2025-07-01", "Security Team", "5 years", "Current"],
        ["EVD-824.2-024", "Tape Encryption Key Escrow Procedure", "Procedure", "6. Backups", "SharePoint/Procedures", "2025-09-15", "Backup Team", "5 years", "Current"],
        ["EVD-824.2-025", "IronKey USB Deployment List", "Inventory", "7. Removable Media", "SharePoint/Inventory/USB", "2026-01-08", "IT Admin", "2 years", "Current"],
        ["EVD-824.2-026", "Removable Media Policy v2.5", "Policy", "7. Removable Media", "SharePoint/Policies/Media", "2025-08-01", "Security Team", "5 years", "Current"],
        ["EVD-824.2-027", "DLP USB Blocking Configuration", "GPO Export", "7. Removable Media", "AD/GPO/DLP", "2025-11-20", "Security Team", "2 years", "Current"],
        ["EVD-824.2-028", "Encrypted USB Vendor Comparison", "Analysis Doc", "7. Removable Media", "SharePoint/Procurement", "2025-10-10", "IT Procurement", "3 years", "Current"],
    ]

    count = safely_write_data(ws, 10, evidence_data)
    logger.info("    Evidence Register: %d evidence documents", count)

    return count


def populate_approval_signoff(wb):
    """Populate Approval Sign-Off."""

    if "Approval Sign-Off" not in wb.sheetnames:
        return 0

    ws = wb["Approval Sign-Off"]

    today = datetime.now()
    ws["B5"] = "Emily Rodriguez"
    ws["B6"] = "Security Analyst"
    ws["B7"] = "Information Security"
    ws["B8"] = "emily.rodriguez@company.com"
    ws["B9"] = today.strftime("%Y-%m-%d")

    ws["B14"] = "David Park"
    ws["B15"] = (today + timedelta(days=2)).strftime("%Y-%m-%d")
    ws["B16"] = "Encryption coverage is comprehensive across storage systems. Legacy database encryption requires attention."

    ws["B21"] = "Michael Chen"
    ws["B22"] = (today + timedelta(days=3)).strftime("%Y-%m-%d")
    ws["B23"] = "Storage encryption controls meet requirements. Priority: Legacy backup system migration."

    ws["B28"] = "Jennifer Williams"
    ws["B29"] = (today + timedelta(days=5)).strftime("%Y-%m-%d")
    ws["B30"] = "Approved with conditions"
    ws["B31"] = "Approved pending Q2 2026 completion of MongoDB KMS integration and legacy backup migration."

    ws["B36"] = (today + timedelta(days=90)).strftime("%Y-%m-%d")
    ws["B37"] = "David Park"
    ws["B38"] = "Q2 2026 reassessment focusing on database encryption standardization"

    logger.info("    Approval Sign-Off: Complete")

    return 1


def main():
    """Main function to populate the Data Storage assessment workbook."""
    try:
        if len(sys.argv) < 2:
            logger.error("Usage: python3 populate_a824_2_data_storage.py <workbook.xlsx>")
            return 1

        filepath = sys.argv[1]

        logger.info("=" * 80)
        logger.info("ISMS-IMP-A.8.24.2 - Data Storage Assessment")
        logger.info("Comprehensive Data Population for CISO Presentation")
        logger.info("=" * 80)

        try:
            wb = load_workbook(filepath)
            logger.info("Loading: %s", filepath)
            logger.info("Sheets found: %d", len(wb.sheetnames))

            logger.info("[1/3] Populating 7 Assessment Sheets...")
            assessment_count = populate_assessment_sheets(wb)

            logger.info("[2/3] Populating Evidence Register...")
            evidence_count = populate_evidence_register(wb)

            logger.info("[3/3] Populating Approval Sign-Off...")
            approval_count = populate_approval_signoff(wb)

            wb.save(filepath)
            logger.info("Saved: %s", filepath)

            logger.info("=" * 80)
            logger.info("DATA POPULATION COMPLETE")
            logger.info("=" * 80)
            logger.info("Summary:")
            logger.info("  - Assessment Entries: %d", assessment_count)
            logger.info("  - Evidence Documents: %d", evidence_count)
            logger.info("  - Total Data Points: %d", assessment_count + evidence_count)
            logger.info("CISO-Ready: Professional storage encryption assessment")
            logger.info("=" * 80)

            return 0

        except Exception as e:
            logger.error("Error: %s", e)
            import traceback
            traceback.print_exc()
            return 1

    except Exception as e:
        logger.error("Population failed with exception: %s", str(e))
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (standardized: license, imports, logging, exit codes)
# QA_TOOL: Claude Code Deep Scan
# STANDARDIZATION: v1.0 - AGPL-3.0/Commercial dual-license
# =============================================================================
