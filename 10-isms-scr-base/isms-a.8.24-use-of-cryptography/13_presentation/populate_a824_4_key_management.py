#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
ISMS-IMP-A.8.24.4 - Key Management Assessment
Comprehensive Data Population for CISO Presentation

Generated: 2026-01-13
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


def safely_write_data(ws, start_row, data):
    """Safely write data to worksheet, handling merged cells."""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except Exception as e:  # Skip merged/protected cells
                continue
        entries_written += 1
    return entries_written


def populate_assessment_sheets(wb):
    """Populate all assessment sheets."""
    sheet_data = {
        "1. Key Generation": [
            ["TLS Certificates (Public)", "Public-facing services", "RSA", "4096 bits", "Let's Encrypt / DigiCert", "Public", "Active", "✅ Compliant", "Certificate inventory", "", "No"],
            ["Database Encryption Keys", "Data-at-rest protection", "AES", "256 bits", "AWS KMS / Azure Key Vault", "Restricted", "Active", "✅ Compliant", "KMS key inventory", "", "No"],
            ["API Signing Keys (JWT)", "API authentication", "ECDSA P-256", "256 bits", "HashiCorp Vault", "Confidential", "Active", "✅ Compliant", "Vault key audit", "", "No"],
            ["SSH Host Keys", "Server authentication", "Ed25519", "256 bits", "Ansible automated", "Internal", "Active", "✅ Compliant", "SSH key inventory", "", "No"],
            ["Code Signing Keys", "Software release signing", "RSA", "4096 bits", "YubiKey HSM", "Restricted", "Active", "✅ Compliant", "Code signing logs", "", "No"],
            ["Symmetric Encryption Keys", "Application encryption", "AES", "256 bits", "Cloud KMS", "Confidential", "Active", "✅ Compliant", "Application key docs", "", "No"],
            ["Legacy DES Keys", "Legacy system (deprecated)", "DES", "56 bits", "File-based", "Confidential", "Deprecated", "❌ Non-Compliant", "Legacy system docs", "Migrate to AES-256", "Yes"],
            ["Key Encryption Keys (KEK)", "Key wrapping", "AES", "256 bits", "HSM-backed", "Restricted", "Active", "✅ Compliant", "HSM audit logs", "", "No"],
            ["Master Encryption Keys", "Top-level keys", "AES", "256 bits", "Hardware HSM", "Restricted", "Active", "✅ Compliant", "HSM key ceremony docs", "", "No"],
        ],

        "2. Key Storage": [
            ["Production TLS Keys", "TLS private keys", "RSA", "4096 bits", "AWS ACM / Azure Key Vault", "Public", "HSM-backed", "✅ Compliant", "ACM configuration", "", "No"],
            ["Database Master Keys", "DB encryption keys", "AES", "256 bits", "Azure Key Vault + HSM", "Restricted", "HSM-backed", "✅ Compliant", "Key Vault access logs", "", "No"],
            ["Application Secrets", "API keys, passwords", "Various", "256+ bits", "HashiCorp Vault", "Confidential", "Encrypted at rest", "✅ Compliant", "Vault access policies", "", "No"],
            ["Kubernetes Secrets", "K8s application secrets", "Various", "Varies", "etcd encrypted", "Internal", "etcd encryption", "✅ Compliant", "K8s encryption config", "", "No"],
            ["Code Signing Keys", "Software signing", "RSA", "4096 bits", "YubiKey HSM", "Restricted", "Hardware device", "✅ Compliant", "HSM access logs", "", "No"],
            ["SSH Private Keys", "User authentication", "Ed25519", "256 bits", "User ~/.ssh/", "Internal", "Filesystem + passphrase", "⚠️ Partial", "Endpoint scan results", "Enforce passphrase policy", "Yes"],
            ["Service Account Keys", "GCP service accounts", "RSA", "2048 bits", "GCP Secret Manager", "Confidential", "Encrypted", "✅ Compliant", "GCP audit logs", "", "No"],
            ["Backup Encryption Keys", "Backup protection", "AES", "256 bits", "Key escrow service", "Restricted", "Secure escrow", "✅ Compliant", "Escrow audit trail", "", "No"],
            ["Development Keys (local)", "Dev environment", "Various", "Varies", "Local filesystem", "Internal", "Unprotected", "❌ Non-Compliant", "Dev key audit", "Implement Vault for dev", "Yes"],
        ],

        "3. Key Rotation": [
            ["TLS Certificates", "Public TLS certs", "RSA", "4096 bits", "Automated (Let's Encrypt)", "Public", "90 days", "✅ Compliant", "Auto-renewal logs", "", "No"],
            ["Database Encryption Keys", "DB master keys", "AES", "256 bits", "Manual rotation process", "Restricted", "365 days", "⚠️ Partial", "KMS rotation schedule", "Automate rotation", "Yes"],
            ["API Keys (Customer)", "Customer API access", "Random", "256 bits", "Vault auto-rotation", "Confidential", "180 days", "✅ Compliant", "Vault rotation logs", "", "No"],
            ["JWT Signing Keys", "Token signing", "ECDSA", "256 bits", "Automated rotation", "Confidential", "180 days", "✅ Compliant", "API gateway logs", "", "No"],
            ["GCP Service Keys", "Service account keys", "RSA", "2048 bits", "Automated", "Confidential", "90 days", "✅ Compliant", "GCP IAM audit", "", "No"],
            ["Azure Storage Keys", "Storage account access", "AES", "256 bits", "Manual", "Confidential", "Never rotated", "❌ Non-Compliant", "Azure Storage audit", "Enable auto-rotation", "Yes"],
            ["SSH Host Keys", "Server host keys", "Ed25519", "256 bits", "Manual", "Internal", "No rotation policy", "⚠️ Partial", "SSH key inventory", "Implement rotation policy", "Yes"],
            ["Encryption Keys (Apps)", "Application-level", "AES", "256 bits", "Automated via Vault", "Confidential", "180 days", "✅ Compliant", "Application key logs", "", "No"],
        ],

        "4. Key Backup & Recovery": [
            ["HSM Master Keys", "Hardware HSM keys", "AES", "256 bits", "M of N key ceremony", "Restricted", "Secure key ceremony", "✅ Compliant", "Key ceremony documentation", "", "No"],
            ["Database Master Keys", "DB encryption keys", "AES", "256 bits", "KMS backup + escrow", "Restricted", "Cloud KMS backup", "✅ Compliant", "KMS backup policies", "", "No"],
            ["BitLocker Recovery Keys", "Disk encryption recovery", "AES", "256 bits", "Active Directory escrow", "Restricted", "AD recovery key store", "✅ Compliant", "BitLocker recovery audit", "", "No"],
            ["FileVault Recovery Keys", "Mac disk encryption", "AES", "256 bits", "MDM escrow", "Restricted", "Jamf Pro escrow", "✅ Compliant", "MDM recovery key audit", "", "No"],
            ["Backup Encryption Keys", "Backup system keys", "AES", "256 bits", "Secure key escrow service", "Restricted", "Iron Mountain escrow", "✅ Compliant", "Escrow service SLA", "", "No"],
            ["Code Signing Keys", "Software signing", "RSA", "4096 bits", "HSM + offline backup", "Restricted", "Secure offline storage", "✅ Compliant", "Backup key audit", "", "No"],
            ["Root CA Keys", "PKI root keys", "RSA", "4096 bits", "Offline HSM + safe", "Restricted", "Air-gapped HSM", "✅ Compliant", "CA key ceremony docs", "", "No"],
            ["Personal SSH Keys", "User SSH keys", "Various", "Varies", "No backup", "Internal", "User responsibility", "⚠️ Partial", "User key management policy", "Centralized key management", "Yes"],
        ],

        "5. Certificate Management": [
            ["Public TLS Certificates", "External services", "RSA/ECDSA", "2048-4096 bits", "Let's Encrypt + DigiCert", "Public", "Active", "✅ Compliant", "Certificate inventory system", "", "No"],
            ["Internal PKI Certificates", "Internal services", "RSA", "2048 bits", "Internal CA", "Internal", "Active", "✅ Compliant", "AD Certificate Services", "", "No"],
            ["Client Authentication Certs", "User/device auth", "RSA/ECDSA", "2048-4096 bits", "Internal PKI + MDM", "Internal", "Active", "✅ Compliant", "PKI audit logs", "", "No"],
            ["Code Signing Certificates", "Software signing", "RSA", "4096 bits", "DigiCert EV", "Restricted", "Active", "✅ Compliant", "Code signing inventory", "", "No"],
            ["S/MIME Certificates", "Email encryption", "RSA", "2048 bits", "Public CA + Internal", "Internal", "Active", "⚠️ Partial", "S/MIME deployment report", "Increase adoption to 100%", "Yes"],
            ["VPN Certificates", "VPN authentication", "RSA/ECDSA", "2048-4096 bits", "Internal PKI", "Internal", "Active", "✅ Compliant", "VPN certificate audit", "", "No"],
            ["IoT Device Certificates", "IoT authentication", "ECDSA", "256 bits", "Internal IoT CA", "Internal", "Active", "✅ Compliant", "IoT PKI documentation", "", "No"],
            ["Certificate Expiry Monitoring", "Proactive monitoring", "N/A", "N/A", "Automated monitoring (30/7 days)", "N/A", "Active", "✅ Compliant", "Monitoring dashboard", "", "No"],
            ["Legacy Self-Signed Certs", "Legacy systems", "RSA", "1024 bits", "Self-signed", "Internal", "Deprecated", "❌ Non-Compliant", "Legacy cert inventory", "Replace with proper PKI", "Yes"],
        ],
    }

    total_written = 0
    for sheet_name, data in sheet_data.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            count = safely_write_data(ws, 8, data)
            total_written += count
            logger.info("    %s: %d entries", sheet_name, count)
    return total_written


def populate_evidence_register(wb):
    """Populate Evidence Register."""
    if "Evidence Register" not in wb.sheetnames:
        return 0
    ws = wb["Evidence Register"]
    # Columns (must match generator): Evidence ID, Assessment Area, Evidence Type,
    #          Description, Location/Path, Date Collected, Collected By, Verification Status
    evidence_data = [
        ["EVD-824.4-001", "1. Key Generation", "Documentation", "Key Generation Standard v2.0", "SharePoint/Standards/Keys", "2025-09-01", "Security Team", "Verified"],
        ["EVD-824.4-002", "1. Key Generation", "Documentation", "AWS KMS Key Policy Documentation", "AWS/KMS/Policies", "2026-01-10", "Cloud Team", "Verified"],
        ["EVD-824.4-003", "1. Key Generation", "Audit log", "HashiCorp Vault Key Audit Report", "Vault/Audit/Q4-2025", "2025-12-20", "Security Team", "Verified"],
        ["EVD-824.4-004", "1. Key Generation", "Documentation", "Code Signing Key Ceremony Documentation", "SharePoint/Ceremonies/Code-Signing", "2025-08-15", "DevOps Team", "Verified"],
        ["EVD-824.4-005", "1. Key Generation", "Documentation", "YubiKey HSM Deployment Guide", "SharePoint/Guides/HSM", "2025-10-01", "Security Team", "Verified"],
        ["EVD-824.4-006", "2. Key Storage", "Screenshot", "AWS ACM Configuration Screenshot", "SharePoint/Evidence/ACM", "2026-01-10", "Cloud Team", "Verified"],
        ["EVD-824.4-007", "2. Key Storage", "Configuration file", "Azure Key Vault Access Policies", "Azure/KeyVault/Policies", "2026-01-10", "Cloud Team", "Verified"],
        ["EVD-824.4-008", "2. Key Storage", "Documentation", "Vault Access Control Policies", "Vault/Policies", "2025-11-15", "Security Team", "Verified"],
        ["EVD-824.4-009", "2. Key Storage", "Configuration file", "Kubernetes etcd Encryption Configuration", "Git/k8s/etcd-encryption", "2025-11-20", "Platform Team", "Verified"],
        ["EVD-824.4-010", "2. Key Storage", "Documentation", "SSH Key Passphrase Enforcement Policy", "SharePoint/Policies/SSH", "2025-09-01", "Security Team", "Verified"],
        ["EVD-824.4-011", "3. Key Rotation", "Audit log", "TLS Certificate Auto-Renewal Logs", "CertBot/Logs/2025", "2026-01-10", "DevOps", "Verified"],
        ["EVD-824.4-012", "3. Key Rotation", "Documentation", "Key Rotation Schedule Q1 2026", "SharePoint/Schedules/Keys", "2026-01-08", "Security Team", "Verified"],
        ["EVD-824.4-013", "3. Key Rotation", "Configuration file", "Vault Key Rotation Configuration", "Vault/Config/Rotation", "2025-12-01", "Security Team", "Verified"],
        ["EVD-824.4-014", "3. Key Rotation", "Documentation", "Azure Storage Key Rotation Remediation Plan", "SharePoint/Projects/Azure-Keys", "2026-01-05", "Cloud Team", "Pending verification"],
        ["EVD-824.4-015", "4. Key Backup & Recovery", "Documentation", "HSM Master Key Ceremony Procedure", "SharePoint/Procedures/HSM", "2025-07-01", "Security Team", "Verified"],
        ["EVD-824.4-016", "4. Key Backup & Recovery", "Compliance report", "BitLocker Recovery Key Escrow Audit", "AD/BitLocker/Audit", "2025-12-15", "Windows Team", "Verified"],
        ["EVD-824.4-017", "4. Key Backup & Recovery", "Screenshot", "Jamf Pro FileVault Key Escrow Config", "Jamf/FileVault/Config", "2026-01-09", "Mac Team", "Verified"],
        ["EVD-824.4-018", "4. Key Backup & Recovery", "Documentation", "Root CA Key Ceremony Documentation", "PKI/Root-CA/Ceremony", "2024-06-15", "Security Team", "Verified"],
        ["EVD-824.4-019", "5. Certificate Management", "Certificate inventory", "Certificate Inventory System Export", "CertManager/Inventory", "2026-01-10", "Security Team", "Verified"],
        ["EVD-824.4-020", "5. Certificate Management", "Documentation", "Internal PKI Architecture Documentation", "SharePoint/Architecture/PKI", "2025-09-20", "Security Architect", "Verified"],
        ["EVD-824.4-021", "5. Certificate Management", "Screenshot", "Certificate Expiry Monitoring Dashboard", "SharePoint/Evidence/Cert-Monitor", "2026-01-09", "Security Team", "Verified"],
        ["EVD-824.4-022", "5. Certificate Management", "Compliance report", "S/MIME Deployment Status Report", "SharePoint/Reports/SMIME", "2026-01-08", "Email Team", "Verified"],
    ]
    count = safely_write_data(ws, 5, evidence_data)  # Start at row 5 (after headers at row 4)
    logger.info("    Evidence Register: %d evidence documents", count)
    return count


def populate_approval_signoff(wb):
    """Populate Approval Sign-Off."""

    if "Approval Sign-Off" not in wb.sheetnames:
        return 0

    ws = wb["Approval Sign-Off"]

    today = datetime.now()

    # Helper function to safely write to cells (handles merged cells)
    def safe_write(cell_ref, value):
        try:
            cell = ws[cell_ref]
            if not isinstance(cell, MergedCell):
                cell.value = value
        except Exception as e:  # Skip merged/protected cells
            pass

    # Assessment completion
    safe_write("B5", "Robert Kim")
    safe_write("B6", "Cryptography Specialist")
    safe_write("B7", "Information Security")
    safe_write("B8", "robert.kim@company.com")
    safe_write("B9", today.strftime("%Y-%m-%d"))

    # Technical Review
    safe_write("B14", "Patricia Lopez")
    safe_write("B15", (today + timedelta(days=2)).strftime("%Y-%m-%d"))
    safe_write("B16", "Key management lifecycle controls are comprehensive. Legacy DES keys and Azure storage key rotation require immediate attention.")

    # Security Review
    safe_write("B21", "Michael Chen")
    safe_write("B22", (today + timedelta(days=3)).strftime("%Y-%m-%d"))
    safe_write("B23", "Strong key management posture. Automated rotation coverage is good. Legacy system migration is priority.")

    # Management Approval
    safe_write("B28", "Jennifer Williams")
    safe_write("B29", (today + timedelta(days=5)).strftime("%Y-%m-%d"))
    safe_write("B30", "Approved with conditions")
    safe_write("B31", "Approved pending Q2 2026 completion of legacy DES key migration and Azure storage key auto-rotation.")

    # Next Review
    safe_write("B36", (today + timedelta(days=90)).strftime("%Y-%m-%d"))
    safe_write("B37", "Patricia Lopez")
    safe_write("B38", "Q2 2026 reassessment focusing on automated key rotation coverage")

    logger.info("    Approval Sign-Off: Complete")
    return 1


def main():
    """Main function to populate the Key Management assessment workbook."""
    try:
        if len(sys.argv) < 2:
            logger.error("Usage: python3 populate_a824_4_key_management.py <workbook.xlsx>")
            return 1

        filepath = sys.argv[1]

        logger.info("=" * 80)
        logger.info("ISMS-IMP-A.8.24.4 - Key Management Assessment")
        logger.info("Comprehensive Data Population for CISO Presentation")
        logger.info("=" * 80)

        try:
            wb = load_workbook(filepath)
            logger.info("Loading: %s", filepath)
            logger.info("Sheets found: %d", len(wb.sheetnames))

            logger.info("[1/3] Populating 5 Assessment Sheets...")
            assessment_count = populate_assessment_sheets(wb)

            logger.info("[2/3] Populating Evidence Register...")
            evidence_count = populate_evidence_register(wb)

            logger.info("[3/3] Populating Approval Sign-Off...")
            approval_count = populate_approval_signoff(wb)

            wb.save(filepath)
            logger.info("Saved: %s", filepath)

            logger.info("=" * 80)
            logger.info("DATA POPULATION COMPLETE")
            logger.info("=" * 80)
            logger.info("Summary:")
            logger.info("  - Assessment Entries: %d", assessment_count)
            logger.info("  - Evidence Documents: %d", evidence_count)
            logger.info("  - Total Data Points: %d", assessment_count + evidence_count)
            logger.info("CISO-Ready: Professional key management assessment")
            logger.info("=" * 80)

            return 0

        except Exception as e:
            logger.error("Error: %s", e)
            import traceback
            traceback.print_exc()
            return 1

    except Exception as e:
        logger.error("Population failed with exception: %s", str(e))
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (standardized: license, imports, logging, exit codes)
# QA_TOOL: Claude Code Deep Scan
# STANDARDIZATION: v1.0 - AGPL-3.0/Commercial dual-license
# =============================================================================
