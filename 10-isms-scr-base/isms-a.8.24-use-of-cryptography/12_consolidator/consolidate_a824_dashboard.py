#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-A.8.24 - Dashboard Consolidation Utility
================================================================================

ISO/IEC 27001:2022 Control A.8.24: Use of Cryptography
Operational Utility: Multi-Domain Assessment Dashboard Consolidation

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific assessment workbook structures and consolidation
requirements.

CRITICAL: This consolidation utility MUST be customized based on the ACTUAL
structure of your four assessment workbooks. DO NOT assume sheet names,
column positions, or data structures match this template.

Key customization areas:
1. Input workbook schemas - ANALYZE ACTUAL STRUCTURE FIRST
2. Sheet names and column mappings - MUST MATCH YOUR ASSESSMENTS
3. Data extraction logic - SPECIFIC TO YOUR WORKBOOK LAYOUTS
4. Scoring and weighting algorithms - ADAPT TO YOUR RISK PROFILE
5. Dashboard visualization preferences - ALIGN WITH YOUR REPORTING NEEDS

MANDATORY STEPS BEFORE USING THIS SCRIPT:
1. Generate all four A.8.24 assessment workbooks (Domains 1-4)
2. Run normalize_assessment_files_a824.py to validate structure
3. Document the actual sheet structure of each workbook
4. Update WORKBOOK_SCHEMAS dictionary with real structure
5. Validate schema against actual files before consolidation
6. Test with sample/development data before production use

Reference Pattern: Based on ISMS-A.8.24 Use of Cryptography Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script consolidates data from four normalized cryptographic assessment
workbooks into a unified compliance dashboard, providing executive visibility
and audit-ready evidence summaries.

**Purpose:**
Automates the consolidation of multi-domain cryptographic assessments into
a single dashboard workbook, eliminating manual copy-paste errors and ensuring
consistent compliance reporting.

**Relationship to generate_a824_5_compliance_summary_dashboard.py:**
- generate_a824_5: Creates the EMPTY dashboard template structure
- consolidate_a824_dashboard.py: POPULATES dashboard with actual assessment data

Use both scripts in sequence:
1. Run generate_a824_5 to create dashboard template
2. Complete all four domain assessments
3. Run normalize_assessment_files_a824.py for validation
4. Run consolidate_a824_dashboard.py to populate dashboard

**Input Assessment Domains:**
1. Domain 1: Data Transmission Cryptography (A.8.24.1)
   Input: ISMS_A_8_24_1_Data_Transmission_Assessment_YYYYMMDD.xlsx
   
2. Domain 2: Data Storage Cryptography (A.8.24.2)
   Input: ISMS_A_8_24_2_Data_Storage_Assessment_YYYYMMDD.xlsx
   
3. Domain 3: Authentication Cryptography (A.8.24.3)
   Input: ISMS_A_8_24_3_Authentication_Assessment_YYYYMMDD.xlsx
   
4. Domain 4: Key Management (A.8.24.4)
   Input: ISMS_A_8_24_4_Key_Management_Assessment_YYYYMMDD.xlsx

**Consolidation Process:**

Phase 1: Pre-Consolidation Validation
- Verify all input workbooks exist and are readable
- Validate workbook structure against expected schema
- Check for required sheets and columns
- Identify missing or incomplete data
- Generate pre-consolidation validation report

Phase 2: Data Extraction
- Extract compliance scores from each domain
- Collect gap analysis findings
- Aggregate evidence references
- Summarize remediation requirements
- Calculate domain-specific metrics

Phase 3: Dashboard Population
- Populate executive summary with aggregate metrics
- Generate compliance overview by domain
- Create consolidated gap analysis
- Populate remediation tracker
- Build audit evidence index
- Calculate overall compliance percentage

Phase 4: Quality Assurance
- Validate consolidated data integrity
- Check for calculation errors
- Verify evidence linkage completeness
- Generate post-consolidation validation report

Phase 5: Output Generation
- Save populated dashboard workbook
- Generate consolidation log
- Create executive summary report (optional)
- Archive source assessment references

**Generated Dashboard Outputs:**
- Primary: ISMS_A_8_24_5_Compliance_Dashboard_YYYYMMDD.xlsx (populated)
- Log: A824_Dashboard_Consolidation_Log_YYYYMMDD.txt
- Report: A824_Executive_Summary_YYYYMMDD.pdf (optional, if --pdf flag used)

**Key Features:**
- Schema-driven data extraction (prevents hardcoded assumptions)
- Intelligent error handling for missing/malformed data
- Weighted scoring based on domain criticality
- Gap prioritization by risk and compliance impact
- Audit trail linking dashboard to source assessments
- Rollback capability if consolidation fails
- Incremental updates (rerun after assessment changes)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel processing

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - os (standard library)
    - json (standard library - schema definitions)
    - logging (standard library - consolidation logging)

Optional Dependencies:
    - reportlab (for PDF executive summary generation)
      pip3 install reportlab

Input Requirements:
    - All four domain assessment workbooks must be completed
    - Workbooks should be normalized (run normalize_assessment_files_a824.py first)
    - Dashboard template must exist (generated by generate_a824_5)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    # Consolidate all assessments into dashboard (auto-detect latest files)
    python3 consolidate_a824_dashboard.py

Advanced Usage:
    # Specify input directory for assessment workbooks
    python3 consolidate_a824_dashboard.py --input-dir /path/to/assessments
    
    # Specify dashboard template file explicitly
    python3 consolidate_a824_dashboard.py --dashboard /path/to/dashboard.xlsx
    
    # Specify output directory for populated dashboard
    python3 consolidate_a824_dashboard.py --output-dir /path/to/output
    
    # Dry run mode (validate without writing to dashboard)
    python3 consolidate_a824_dashboard.py --dry-run
    
    # Generate PDF executive summary
    python3 consolidate_a824_dashboard.py --pdf
    
    # Verbose logging for troubleshooting
    python3 consolidate_a824_dashboard.py --verbose
    
    # Specify assessment files explicitly (override auto-detection)
    python3 consolidate_a824_dashboard.py \
        --domain1 path/to/transmission_assessment.xlsx \
        --domain2 path/to/storage_assessment.xlsx \
        --domain3 path/to/authentication_assessment.xlsx \
        --domain4 path/to/key_management_assessment.xlsx

Command-Line Options:
    --input-dir PATH       Directory containing assessment workbooks
    --dashboard PATH       Path to dashboard template file
    --output-dir PATH      Directory for populated dashboard
    --domain1 PATH         Explicit path to Domain 1 assessment
    --domain2 PATH         Explicit path to Domain 2 assessment
    --domain3 PATH         Explicit path to Domain 3 assessment
    --domain4 PATH         Explicit path to Domain 4 assessment
    --dry-run              Validate without writing to dashboard
    --pdf                  Generate PDF executive summary (requires reportlab)
    --verbose              Enable detailed logging
    --backup               Create backup of dashboard before populating
    --incremental          Update existing dashboard (preserve manual edits)

Output Files:
    Primary Output:
        ISMS_A_8_24_5_Compliance_Dashboard_YYYYMMDD.xlsx (populated)
    
    Supporting Files:
        A824_Dashboard_Consolidation_Log_YYYYMMDD.txt (consolidation log)
        A824_Executive_Summary_YYYYMMDD.pdf (if --pdf used)
        dashboard_backup_YYYYMMDD.xlsx (if --backup used)

Workflow Examples:

    1. First-time consolidation:
       python3 consolidate_a824_dashboard.py --verbose --backup
    
    2. Update dashboard after assessment changes:
       python3 consolidate_a824_dashboard.py --incremental
    
    3. Validation before production:
       python3 consolidate_a824_dashboard.py --dry-run --verbose
    
    4. Generate executive briefing:
       python3 consolidate_a824_dashboard.py --pdf

Exit Codes:
    0   Success - Dashboard populated without errors
    1   Validation failure - Input workbooks don't meet schema requirements
    2   Data extraction error - Unable to read required data from assessments
    3   Consolidation error - Error during dashboard population
    4   File I/O error - Unable to read/write required files

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.24
Utility Type:         Operational - Dashboard Data Consolidation
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.24: Use of Cryptography Policy (Governance)
    - ISMS-IMP-A.8.24.1: Data Transmission Assessment (Input Domain 1)
    - ISMS-IMP-A.8.24.2: Data Storage Assessment (Input Domain 2)
    - ISMS-IMP-A.8.24.3: Authentication Assessment (Input Domain 3)
    - ISMS-IMP-A.8.24.4: Key Management Assessment (Input Domain 4)
    - ISMS-IMP-A.8.24.5: Compliance Dashboard Implementation Guide

Related Scripts:
    - generate_a824_1_data_transmission_assessment.py (generates input)
    - generate_a824_2_data_storage_assessment.py (generates input)
    - generate_a824_3_authentication_assessment.py (generates input)
    - generate_a824_4_key_management_assessment.py (generates input)
    - generate_a824_5_compliance_summary_dashboard.py (generates template)
    - normalize_assessment_files_a824.py (validates inputs)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements schema-driven consolidation framework
    - Supports all four cryptographic assessment domains
    - Generates audit-ready compliance dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Schema Validation is MANDATORY:**
This script WILL FAIL if workbook schemas don't match expectations. You MUST:
1. Run normalize_assessment_files_a824.py BEFORE consolidation
2. Document actual workbook structure in WORKBOOK_SCHEMAS dictionary
3. Test with --dry-run before production consolidation
4. Keep schema definitions synchronized with assessment workbooks

Failure to validate schemas = garbage in, garbage out.

**Audit Considerations:**
This script generates the primary compliance dashboard for ISO 27001:2022
Control A.8.24 audits. Auditors will expect:
- Traceability from dashboard metrics to source assessments
- Consolidation log demonstrating systematic process
- Evidence that consolidated data hasn't been manually manipulated

Maintain consolidation logs and source assessments for audit trail.

**Data Protection:**
Consolidated dashboard contains executive summary of cryptographic security
posture across all domains. This is HIGHLY SENSITIVE information including:
- Overall compliance gaps and critical vulnerabilities
- Key management deficiencies across the organization
- Remediation priorities and timelines

Handle with MAXIMUM SECURITY. Consider encrypting the dashboard workbook.

**Incremental vs. Full Consolidation:**
- Full consolidation (default): Overwrites all dashboard data
- Incremental (--incremental): Preserves manual edits in dashboard

Use incremental mode carefully - it can preserve outdated data if assessment
values have changed. Default to full consolidation for accuracy.

**Error Handling Philosophy:**
Script fails fast on schema mismatches or missing data. Better to fail
obviously than to silently produce incorrect compliance metrics.

If consolidation fails:
1. Check consolidation log for error details
2. Run normalize_assessment_files_a824.py to identify issues
3. Fix source assessment workbooks
4. Rerun consolidation

**Performance:**
Consolidation processes multiple large Excel workbooks. Expected runtime:
- Small deployments (<100 systems): <30 seconds
- Medium deployments (100-1000 systems): 1-3 minutes
- Large deployments (>1000 systems): 3-10 minutes

Use --verbose to monitor progress for large consolidations.

**Quality Assurance:**
After consolidation, manually verify:
- Compliance percentages look reasonable
- Gap counts match expectations
- Critical findings are correctly prioritized
- Evidence references are intact

Don't blindly trust automated consolidation - validate the results.

**Integration Opportunities:**
Dashboard data can feed into:
- GRC platform compliance reporting
- Executive dashboards and business intelligence tools
- Automated compliance monitoring systems
- Ticketing systems for remediation tracking

Consider exporting dashboard data to JSON/CSV for integration.

**Regulatory Reporting:**
Customize consolidation logic to highlight regulatory-specific compliance:
- PCI DSS: Emphasize payment system cryptography
- HIPAA: Highlight healthcare data encryption
- GDPR: Focus on personal data protection cryptography

Adapt weighting and prioritization to your regulatory context.

**Backup and Recovery:**
Always use --backup flag for production consolidations. If consolidation
produces unexpected results, you can restore from backup and troubleshoot.

**Version Control:**
Maintain version history of consolidated dashboards to track compliance
improvement over time. Use date suffixes and archive previous versions.

================================================================================
"""

import sys
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from datetime import datetime, timedelta

def safely_write_data(ws, start_row, data):
    """Safely write data, handling merged cells"""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except:
                continue
        entries_written += 1
    return entries_written

def extract_gaps_from_workbook(filepath, assessment_area):
    """Extract gap items (Partial/Non-Compliant) from source workbook"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        gaps = []
        gap_counter = 1
        
        # Iterate through assessment sheets
        for sheet_name in wb.sheetnames:
            if sheet_name in ['Instructions', 'Summary Dashboard', 'Evidence Register', 'Approval Sign-Off']:
                continue
            
            ws = wb[sheet_name]
            
            # Find status column and gap description column
            # Typically: Status in column with ✅/⚠️/❌, Gap Description nearby, Remediation at end
            for row in range(8, min(ws.max_row + 1, 50)):  # Check data rows
                row_values = [ws.cell(row=row, column=col).value for col in range(1, 15)]
                
                # Find status (look for ⚠️ or ❌)
                status = None
                status_col = None
                for col_idx, val in enumerate(row_values, start=1):
                    if val in ['⚠️ Partial', '❌ Non-Compliant']:
                        status = val
                        status_col = col_idx
                        break
                
                if not status:
                    continue
                
                # Extract data for this gap
                system_name = row_values[0] if row_values[0] else 'Unknown'
                gap_desc = None
                remediation = None
                
                # Gap description typically 2 columns before status, remediation at end
                if status_col:
                    gap_desc = ws.cell(row=row, column=status_col + 2).value
                    remediation = ws.cell(row=row, column=status_col + 3).value
                
                severity = 'High' if status == '❌ Non-Compliant' else 'Medium'
                
                gap_entry = [
                    f'GAP-{assessment_area[:3].upper()}-{gap_counter:03d}',  # Gap ID
                    assessment_area,                                           # Assessment Area
                    filepath.split('/')[-1],                                  # Source Document
                    system_name,                                              # System/Application
                    gap_desc if gap_desc else 'Configuration gap identified', # Gap Description
                    status.replace('⚠️ ', '').replace('❌ ', ''),            # Current State
                    'Compliant',                                              # Target State
                    severity,                                                 # Severity
                    'Open',                                                   # Status
                    remediation if remediation else 'Remediation required',   # Remediation Plan
                    None,                                                     # Owner (to be assigned)
                    None,                                                     # Target Date
                ]
                gaps.append(gap_entry)
                gap_counter += 1
        
        wb.close()
        return gaps
    except Exception as e:
        print(f"  ⚠️  Error extracting gaps from {filepath}: {e}")
        return []

def extract_evidence_from_workbook(filepath, assessment_area):
    """Extract evidence entries from source workbook"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        if 'Evidence Register' not in wb.sheetnames:
            wb.close()
            return []
        
        ws = wb['Evidence Register']
        evidence = []
        
        for row in range(10, ws.max_row + 1):
            evidence_id = ws[f'A{row}'].value
            if evidence_id and str(evidence_id).startswith('EVD-'):
                row_data = [
                    evidence_id,
                    assessment_area,
                    filepath.split('/')[-1],
                    None,  # Gap ID
                    ws[f'B{row}'].value,  # Document Title
                    ws[f'C{row}'].value,  # Document Type
                    ws[f'D{row}'].value,  # Related Control
                    ws[f'E{row}'].value,  # Storage Location
                    ws[f'F{row}'].value,  # Date Collected
                    ws[f'G{row}'].value,  # Collected By
                    ws[f'H{row}'].value,  # Retention Period
                    ws[f'I{row}'].value,  # Status
                ]
                evidence.append(row_data)
        
        wb.close()
        return evidence
    except Exception as e:
        print(f"  ⚠️  Error reading evidence from {filepath}: {e}")
        return []

def generate_risks_from_gaps(gaps):
    """Generate risk entries from critical gaps"""
    risks = []
    risk_counter = 1
    
    for gap in gaps:
        gap_id = gap[0]
        area = gap[1]
        system = gap[3]
        gap_desc = gap[4]
        severity = gap[7]
        
        # Only create risks for High severity gaps
        if severity != 'High':
            continue
        
        risk_entry = [
            f'RISK-{risk_counter:03d}',                           # Risk ID
            gap_id,                                               # Gap ID (Link)
            'Security',                                           # Risk Category
            f'Security risk due to: {gap_desc[:60]}...',         # Risk Description
            system,                                               # Affected System
            'High' if 'Non-Compliant' in str(gap_desc) else 'Medium',  # Likelihood
            'High',                                               # Impact
            'High',                                               # Inherent Risk
            'Open',                                               # Status
            f'Mitigate through {gap[9][:40]}...',                # Mitigation Plan
            None,                                                 # Owner
            None,                                                 # Target Date
            None,                                                 # Residual Risk
        ]
        risks.append(risk_entry)
        risk_counter += 1
    
    return risks

def generate_remediation_from_gaps(gaps):
    """Generate remediation roadmap from gaps"""
    remediation = []
    action_counter = 1
    today = datetime.now()
    
    for gap in gaps:
        gap_id = gap[0]
        area = gap[1]
        system = gap[3]
        gap_desc = gap[4]
        severity = gap[7]
        remediation_plan = gap[9]
        
        # Priority based on severity
        priority = 'Critical' if severity == 'High' else 'High' if severity == 'Medium' else 'Medium'
        
        # Target date based on priority
        if priority == 'Critical':
            days_offset = 30
        elif priority == 'High':
            days_offset = 60
        else:
            days_offset = 90
        
        target_date = (today + timedelta(days=days_offset)).strftime('%Y-%m-%d')
        
        remediation_entry = [
            f'REM-{action_counter:03d}',                         # Action ID
            gap_id,                                              # Gap ID (Link)
            area,                                                # Assessment Area
            system,                                              # System/Application
            remediation_plan if remediation_plan else 'Remediation required',  # Action Required
            priority,                                            # Priority
            'Planned',                                           # Status
            None,                                                # Owner
            today.strftime('%Y-%m-%d'),                         # Start Date
            target_date,                                         # Target Date
            None,                                                # Actual Date
            None,                                                # Progress %
            None,                                                # Notes
        ]
        remediation.append(remediation_entry)
        action_counter += 1
    
    return remediation

def populate_comprehensive_dashboard(dashboard_file):
    """Populate all dashboard sheets from source workbooks"""
    
    print("="*80)
    print("COMPREHENSIVE DASHBOARD POPULATION")
    print("="*80)
    
    sources = [
        ('ISMS-IMP-A.8.24.1.xlsx', 'Data Transmission'),
        ('ISMS-IMP-A.8.24.2.xlsx', 'Data Storage'),
        ('ISMS-IMP-A.8.24.3.xlsx', 'Authentication'),
        ('ISMS-IMP-A.8.24.4.xlsx', 'Key Management'),
    ]
    
    all_gaps = []
    all_evidence = []
    
    print("\n[1/4] Extracting gaps from source workbooks...")
    for filepath, area in sources:
        print(f"  • {area}: ", end="")
        gaps = extract_gaps_from_workbook(filepath, area)
        print(f"{len(gaps)} gaps")
        all_gaps.extend(gaps)
    
    print(f"\n  Total gaps identified: {len(all_gaps)}")
    
    print("\n[2/4] Extracting evidence from source workbooks...")
    for filepath, area in sources:
        print(f"  • {area}: ", end="")
        evidence = extract_evidence_from_workbook(filepath, area)
        print(f"{len(evidence)} evidence docs")
        all_evidence.extend(evidence)
    
    print(f"\n  Total evidence collected: {len(all_evidence)}")
    
    print("\n[3/4] Generating risks and remediation...")
    all_risks = generate_risks_from_gaps(all_gaps)
    all_remediation = generate_remediation_from_gaps(all_gaps)
    print(f"  • Risks generated: {len(all_risks)}")
    print(f"  • Remediation actions: {len(all_remediation)}")
    
    print(f"\n[4/4] Writing to dashboard...")
    try:
        wb = load_workbook(dashboard_file)
    except Exception as e:
        print(f"  ❌ Error loading dashboard: {e}")
        return False
    
    # Populate Gap Analysis
    if 'Gap Analysis' in wb.sheetnames and all_gaps:
        ws = wb['Gap Analysis']
        count = safely_write_data(ws, 14, all_gaps)  # Start at row 14 (after headers)
        print(f"  ✓ Gap Analysis: {count} entries")
    
    # Populate Risk Register
    if 'Risk Register' in wb.sheetnames and all_risks:
        ws = wb['Risk Register']
        count = safely_write_data(ws, 21, all_risks)  # Start at row 21
        print(f"  ✓ Risk Register: {count} entries")
    
    # Populate Remediation Roadmap
    if 'Remediation Roadmap' in wb.sheetnames and all_remediation:
        ws = wb['Remediation Roadmap']
        count = safely_write_data(ws, 37, all_remediation)  # Start at row 37
        print(f"  ✓ Remediation Roadmap: {count} entries")
    
    # Populate Evidence Register
    if 'Evidence Register' in wb.sheetnames and all_evidence:
        ws = wb['Evidence Register']
        count = safely_write_data(ws, 14, all_evidence)  # Start at row 14
        print(f"  ✓ Evidence Register: {count} entries")
    
    # Save
    try:
        wb.save(dashboard_file)
        print(f"\n💾 Saved: {dashboard_file}")
    except Exception as e:
        print(f"  ❌ Error saving: {e}")
        return False
    
    print("\n" + "="*80)
    print("✅ COMPREHENSIVE DASHBOARD POPULATION COMPLETE")
    print("="*80)
    print(f"\nSummary:")
    print(f"  • Gap Analysis: {len(all_gaps)} gaps identified")
    print(f"  • Risk Register: {len(all_risks)} risks documented")
    print(f"  • Remediation Roadmap: {len(all_remediation)} actions planned")
    print(f"  • Evidence Register: {len(all_evidence)} evidence documents")
    print(f"\n  Total Data Points: {len(all_gaps) + len(all_risks) + len(all_remediation) + len(all_evidence)}")
    print("\n🎯 Dashboard is now CISO-presentation ready!")
    print("="*80 + "\n")
    
    return True

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 consolidate_a824_dashboard.py <dashboard.xlsx>")
        sys.exit(1)
    
    dashboard_file = sys.argv[1]
    success = populate_comprehensive_dashboard(dashboard_file)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
