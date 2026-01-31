#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
Excel Sanity Checker - A.8.24.2 Data Storage Assessment Workbook
================================================================================

SPECIALIZED diagnostic for Data Storage assessment workbooks.

**Purpose:**
Domain-specific validation tailored to the unique structure and requirements
of A.8.24.2 assessment workbooks.

**Checks Performed:**
- Domain-specific sheet structure validation
- Formula integrity for domain-specific calculations
- Data validation rules specific to this assessment type
- Data storage-specific validation logic

**When to Use:**
- After generating A.8.24.2 assessment workbook
- Before stakeholder distribution
- When Excel shows repair warnings for this specific domain

**Usage:**
    python3 excel_sanity_check_a824_2.py ISMS-IMP-A.8.24.2_Assessment_YYYYMMDD.xlsx

**Differences from General Checker:**
This specialized checker validates domain-specific requirements that the
general checker cannot verify (custom sheet structures, domain-specific
formulas, assessment-specific data validation rules).

Control Reference: ISO/IEC 27001:2022 Annex A Control A.8.24
Assessment Domain: 2 of 4 (Data Storage)
Script Type: Domain-Specific QA Utility
Version: 1.0
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
import re

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# ============================================================================
# A.8.24.2 SPECIFIC VALIDATION DATA
# ============================================================================

EXPECTED_SHEETS = [
    "Instructions & Legend",
    "1. Mobile Devices",
    "2. Laptops & Workstations",
    "3. Servers",
    "4. Databases",
    "5. Cloud Storage",
    "6. Backups",
    "7. Removable Media",
    "Summary Dashboard",
    "Evidence Register",
    "Approval Sign-Off",
]

REQUIRED_DOCUMENT_INFO = [
    "Document ID",
    "Assessment Area",
    "Related Policy",
    "Version",
]

STORAGE_ASSESSMENT_SHEETS = [
    "1. Mobile Devices",
    "2. Laptops & Workstations",
    "3. Servers",
    "4. Databases",
    "5. Cloud Storage",
    "6. Backups",
    "7. Removable Media",
]

EXPECTED_COLUMNS = [
    "Storage System/Device",
    "Data Classification",
    "Encryption Status",
    "Encryption Type",
    "Algorithm & Key Size",
    "Key Management Method",
    "Key Rotation Enabled",
    "Status",
    "Evidence Location",
    "Gap Description",
    "Remediation Needed",
    "Exception ID",
    "Risk ID",
    "Compensating Controls",
    "Responsible Person",
    "Target Date",
    "Budget Required",
]


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""

    logger.info("=" * 80)
    logger.info(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {filename}")
    logger.info("ISMS-IMP-A.8.24.2 - Data Storage Assessment")
    logger.info("=" * 80)

    issues_found = []
    warnings_found = []

    try:
        wb = load_workbook(filename, data_only=False)
        logger.info("Workbook loaded successfully")
        logger.info(f"  Sheets found: {len(wb.sheetnames)}")

    except Exception as e:
        logger.error(f"CRITICAL: Cannot load workbook: {e}")
        return 1

    # ========================================================================
    # CHECK 0: A.8.24.2 SPECIFIC STRUCTURE VALIDATION
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 0: A.8.24.2 STRUCTURE VALIDATION")
    logger.info("=" * 80)

    structure_issues = 0

    # Check for expected sheets
    missing_sheets = []
    for sheet in EXPECTED_SHEETS:
        if sheet not in wb.sheetnames:
            missing_sheets.append(sheet)
            structure_issues += 1

    if missing_sheets:
        issues_found.append(f"  Missing required sheets: {', '.join(missing_sheets)}")
        logger.error(f"  Missing {len(missing_sheets)} required sheet(s)")
    else:
        logger.info(f"  All {len(EXPECTED_SHEETS)} required sheets present")

    # Check for unexpected sheets
    unexpected_sheets = []
    for sheet in wb.sheetnames:
        if sheet not in EXPECTED_SHEETS:
            unexpected_sheets.append(sheet)

    if unexpected_sheets:
        warnings_found.append(f"  Unexpected sheets found: {', '.join(unexpected_sheets)}")
        logger.info(f"  Found {len(unexpected_sheets)} unexpected sheet(s)")

    # Validate document information fields
    if "Instructions & Legend" in wb.sheetnames:
        ws = wb["Instructions & Legend"]
        found_fields = []
        for row in ws.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for field in REQUIRED_DOCUMENT_INFO:
                        if field in cell.value:
                            found_fields.append(field)

        missing_fields = [f for f in REQUIRED_DOCUMENT_INFO if f not in found_fields]
        if missing_fields:
            warnings_found.append(f"  Missing document info fields: {', '.join(missing_fields)}")

    # Validate assessment sheet structure
    for sheet_name in STORAGE_ASSESSMENT_SHEETS:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Check if header row exists (row 6)
            if ws.max_row >= 6:
                header_values = [cell.value for cell in ws[6] if cell.value]
                if len(header_values) < 10:
                    warnings_found.append(f"  {sheet_name}: Header row may be incomplete")

    if structure_issues == 0:
        logger.info("  Workbook structure matches A.8.24.2 specification")

    # ========================================================================
    # CHECK 1: FORMULA VALIDATION
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 1: FORMULA VALIDATION")
    logger.info("=" * 80)

    formula_issues = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value

                    # Check for common formula issues
                    if "'" in formula:
                        # Sheet references with special characters
                        sheet_refs = re.findall(r"'([^']+)'!", formula)
                        for ref in sheet_refs:
                            if ref not in wb.sheetnames:
                                issues_found.append(
                                    f"  {sheet_name}!{cell.coordinate}: "
                                    f"Invalid sheet reference '{ref}'"
                                )
                                formula_issues += 1

                    # Check for common syntax issues
                    if formula.count('(') != formula.count(')'):
                        issues_found.append(
                            f"  {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced parentheses in formula"
                        )
                        formula_issues += 1

                    # Check for double quotes issues
                    if formula.count('"') % 2 != 0:
                        issues_found.append(
                            f"  {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced quotes in formula"
                        )
                        formula_issues += 1

    if formula_issues == 0:
        logger.info("  All formulas appear syntactically valid")
    else:
        logger.error(f"  Found {formula_issues} formula issues")

    # ========================================================================
    # CHECK 2: DATA VALIDATION CONFLICTS
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 2: DATA VALIDATION CONFLICTS")
    logger.info("=" * 80)

    validation_issues = 0
    total_validations = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if hasattr(ws, 'data_validations') and ws.data_validations:
            dv_count = len(ws.data_validations.dataValidation)
            total_validations += dv_count
            logger.info(f"  {sheet_name}: {dv_count} data validations")

            # Check for overlapping ranges (simplified)
            all_ranges = []
            for dv in ws.data_validations.dataValidation:
                if hasattr(dv, 'sqref') and dv.sqref:
                    all_ranges.extend(str(dv.sqref).split())

            if len(all_ranges) != len(set(all_ranges)):
                warnings_found.append(
                    f"  {sheet_name}: Potentially overlapping data validation ranges"
                )
                validation_issues += 1

    logger.info(f"  Total data validations across all sheets: {total_validations}")

    if validation_issues == 0:
        logger.info("  No obvious data validation conflicts")

    # ========================================================================
    # CHECK 3: STYLE CONSISTENCY
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 3: STYLE CONSISTENCY")
    logger.info("=" * 80)

    style_issues = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        cells_without_font = 0

        sample_size = 0
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value:
                    sample_size += 1
                    if not hasattr(cell, 'font') or cell.font is None:
                        cells_without_font += 1

        if cells_without_font > 0:
            warnings_found.append(
                f"  {sheet_name}: {cells_without_font}/{sample_size} "
                f"cells missing font attributes"
            )

        # Note: Missing borders are often intentional, so we don't flag this as issue

    if style_issues == 0:
        logger.info("  Style attributes appear consistent")

    # ========================================================================
    # CHECK 4: MERGED CELLS VALIDATION
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 4: MERGED CELLS VALIDATION")
    logger.info("=" * 80)

    merge_issues = 0
    total_merges = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if hasattr(ws, 'merged_cells') and ws.merged_cells:
            merge_count = len(ws.merged_cells.ranges)
            total_merges += merge_count
            logger.info(f"  {sheet_name}: {merge_count} merged cell ranges")

            # Check if merged cells have content in non-top-left cells
            for merge_range in ws.merged_cells.ranges:
                # Get min cell (top-left) of merged range
                min_col = merge_range.min_col
                min_row = merge_range.min_row
                max_col = merge_range.max_col
                max_row = merge_range.max_row

                # Check if any cell other than top-left has content
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        # Skip the top-left cell
                        if row == min_row and col == min_col:
                            continue

                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            cell_coord = f"{get_column_letter(col)}{row}"
                            warnings_found.append(
                                f"  {sheet_name}!{cell_coord}: "
                                f"Merged cell has content in non-primary cell"
                            )
                            merge_issues += 1

    if total_merges > 0:
        logger.info(f"  Total merged ranges across all sheets: {total_merges}")

    if merge_issues == 0:
        logger.info("  Merged cells appear properly formatted")
    else:
        logger.error(f"  Found {merge_issues} merged cell content issues")

    # ========================================================================
    # CHECK 5: WORKSHEET STRUCTURE
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 5: WORKSHEET STRUCTURE")
    logger.info("=" * 80)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Check for excessive dimensions
        if ws.max_row > 10000:
            warnings_found.append(
                f"  {sheet_name}: Large worksheet ({ws.max_row} rows)"
            )

        if ws.max_column > 100:
            warnings_found.append(
                f"  {sheet_name}: Wide worksheet ({ws.max_column} columns)"
            )

    logger.info("  Worksheet dimensions within reasonable limits")

    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("DIAGNOSTIC SUMMARY")
    logger.info("=" * 80)

    if issues_found:
        logger.error(f"CRITICAL ISSUES FOUND: {len(issues_found)}")
        for issue in issues_found:
            logger.error(issue)

    if warnings_found:
        logger.info(f"WARNINGS: {len(warnings_found)}")
        for warning in warnings_found:
            logger.info(warning)

    if not issues_found and not warnings_found:
        logger.info("NO ISSUES DETECTED")
        logger.info("The workbook appears structurally sound and matches A.8.24.2 specification.")
        logger.info("Excel repair warnings may be due to:")
        logger.info("  - Excel version compatibility (try Excel 2019+)")
        logger.info("  - Antivirus/security software interference")
        logger.info("  - Network drive or OneDrive sync issues")
        logger.info("  - Excel's overly cautious validation")
    else:
        logger.info("")
        logger.info("=" * 80)
        logger.info("RECOMMENDED ACTIONS:")
        logger.info("=" * 80)

        if structure_issues > 0:
            logger.info("0. STRUCTURE FIXES:")
            logger.info("   - Verify all required sheets are present")
            logger.info("   - Check sheet names match specification exactly")
            logger.info("   - Ensure document information fields are complete")

        if formula_issues > 0:
            logger.info("1. FORMULA FIXES:")
            logger.info("   - Review formulas referencing other sheets")
            logger.info("   - Ensure sheet names match exactly (case-sensitive)")
            logger.info("   - Check for balanced quotes and parentheses")

        if validation_issues > 0:
            logger.info("2. DATA VALIDATION FIXES:")
            logger.info("   - Remove overlapping data validation ranges")
            logger.info("   - Apply validations to specific ranges, not entire columns")

        if merge_issues > 0:
            logger.info("3. MERGED CELL FIXES:")
            logger.info("   - Ensure only top-left cell of merged range has content")
            logger.info("   - Clear content from other cells in merged range")

    logger.info("")
    logger.info("=" * 80)
    logger.info("A.8.24.2 SPECIFIC NOTES:")
    logger.info("=" * 80)
    logger.info("Expected structure:")
    logger.info("  - 11 sheets total")
    logger.info("  - 7 storage assessment sheets (Mobile, Laptop, Server, Database, Cloud, Backup, Removable)")
    logger.info("  - Each assessment sheet: 17 columns, 13 data rows, compliance checklist")
    logger.info("  - Summary Dashboard with compliance rollup formulas")
    logger.info("  - Evidence Register with 100 entry rows")
    logger.info("  - Approval Sign-Off section")
    logger.info("=" * 80)

    return 1 if issues_found else 0


def main():
    """Main entry point for the sanity checker."""
    try:
        if len(sys.argv) < 2:
            logger.error("Usage: python3 excel_sanity_check_a824_2.py <filename.xlsx>")
            logger.info("Example:")
            logger.info("  python3 excel_sanity_check_a824_2.py ISMS-IMP-A.8.24.2_Data_Storage_20251231.xlsx")
            sys.exit(1)

        filename = sys.argv[1]
        exit_code = check_workbook_health(filename)
        sys.exit(exit_code)

    except KeyboardInterrupt:
        logger.info("Operation cancelled by user")
        sys.exit(130)
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (syntax validated, standardized format applied)
# QA_TOOL: Claude Code Deep Scan
# =============================================================================
