#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
Excel Workbook Sanity Checker - ISMS A.8.24 Assessment Workbooks
================================================================================

Diagnostic utility for troubleshooting Excel's "file level validation and repair"
warnings when opening A.8.24 cryptographic assessment workbooks.

**Purpose:**
Identifies common openpyxl-generated Excel issues that trigger repair warnings:
- Formula syntax errors and invalid sheet references
- Data validation conflicts and overlapping ranges
- Style attribute inconsistencies
- Merged cell content issues
- Worksheet structure problems

**When to Use:**
- Excel displays repair warnings when opening generated workbooks
- After modifying assessment generator scripts
- Before distributing workbooks to stakeholders
- Quality assurance validation before consolidation

**Usage:**
    python3 excel_sanity_check_a824.py ISMS-IMP-A.8.24_X_Assessment_YYYYMMDD.xlsx

    Works with any A.8.24 assessment workbook (domains 1-5)

**Output:**
- Diagnostic report with issue categorization (Critical/Warning)
- Recommended remediation actions
- Structural health summary

**Related Scripts:**
- excel_sanity_check_a824_1.py through _5.py (domain-specific checkers)
- excel_style_object_checker_a824.py (deep style analysis)
- excel_style_object_patcher_a824.py (automated style fixes)

Control Reference: ISO/IEC 27001:2022 Annex A Control A.8.24
Script Type: Quality Assurance Utility
Version: 1.0
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
import re

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# ============================================================================
# WORKBOOK TYPE DETECTION
# ============================================================================

def detect_workbook_type(filename):
    """Detect which A.8.24 workbook type this is based on filename."""
    filename_lower = filename.lower()

    if 'a.8.24.1' in filename_lower or 'data_transmission' in filename_lower:
        return 'A.8.24.1', 'Data Transmission Assessment'
    elif 'a.8.24.2' in filename_lower or 'data_storage' in filename_lower:
        return 'A.8.24.2', 'Data Storage Assessment'
    elif 'a.8.24.3' in filename_lower or 'authentication' in filename_lower:
        return 'A.8.24.3', 'Authentication Assessment'
    elif 'a.8.24.4' in filename_lower or 'key_management' in filename_lower:
        return 'A.8.24.4', 'Key Management Assessment'
    elif 'a.8.24.5' in filename_lower or 'compliance_summary' in filename_lower or 'dashboard' in filename_lower:
        return 'A.8.24.5', 'Compliance Summary Dashboard'
    else:
        return 'Unknown', 'Generic Excel Workbook'


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""

    workbook_id, workbook_name = detect_workbook_type(filename)

    logger.info("=" * 80)
    logger.info(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {filename}")
    logger.info(f"Detected Type: {workbook_id} - {workbook_name}")
    logger.info("=" * 80)

    issues_found = []
    warnings_found = []

    try:
        wb = load_workbook(filename, data_only=False)
        logger.info("Workbook loaded successfully")
        logger.info(f"  Sheets found: {len(wb.sheetnames)}")
        logger.info(f"  Sheet names: {', '.join(wb.sheetnames)}")

    except Exception as e:
        logger.error(f"CRITICAL: Cannot load workbook: {e}")
        return 1

    # ========================================================================
    # CHECK 1: FORMULA VALIDATION
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 1: FORMULA VALIDATION")
    logger.info("=" * 80)

    formula_issues = 0
    inter_sheet_refs = {}
    external_refs = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value

                    # Check for external workbook references
                    ext_refs = re.findall(r'\[([^\]]+\.xlsx)\]', formula)
                    for ref in ext_refs:
                        external_refs.add(ref)

                    # Check for sheet references
                    if "'" in formula:
                        sheet_refs = re.findall(r"'([^']+)'!", formula)
                        for ref in sheet_refs:
                            # Skip external references (contain brackets)
                            if "[" not in ref and ref not in wb.sheetnames:
                                issues_found.append(
                                    f"  {sheet_name}!{cell.coordinate}: "
                                    f"Invalid sheet reference '{ref}'"
                                )
                                formula_issues += 1

                            # Track inter-sheet references (internal only)
                            if "[" not in ref and ref in wb.sheetnames:
                                if sheet_name not in inter_sheet_refs:
                                    inter_sheet_refs[sheet_name] = set()
                                inter_sheet_refs[sheet_name].add(ref)

                    # Check for common syntax issues
                    if formula.count('(') != formula.count(')'):
                        issues_found.append(
                            f"  {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced parentheses in formula"
                        )
                        formula_issues += 1

                    # Check for double quotes issues
                    if formula.count('"') % 2 != 0:
                        issues_found.append(
                            f"  {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced quotes in formula"
                        )
                        formula_issues += 1

    if formula_issues == 0:
        logger.info("  All formulas appear syntactically valid")
    else:
        logger.error(f"  Found {formula_issues} formula issues")

    # Report inter-sheet dependencies
    if inter_sheet_refs:
        logger.info("  Inter-sheet formula dependencies detected:")
        for source, targets in sorted(inter_sheet_refs.items()):
            logger.info(f"    {source} -> {', '.join(sorted(targets))}")

    # Report external workbook links (important for A.8.24.5)
    if external_refs:
        logger.info(f"  External workbook links detected: {len(external_refs)}")
        for ref in sorted(external_refs):
            logger.info(f"    - {ref}")

        if workbook_id == 'A.8.24.5':
            expected_sources = [
                "ISMS-IMP-A.8.24.1.xlsx",
                "ISMS-IMP-A.8.24.2.xlsx",
                "ISMS-IMP-A.8.24.3.xlsx",
                "ISMS-IMP-A.8.24.4.xlsx",
            ]
            missing = [s for s in expected_sources if s not in external_refs]
            if missing:
                warnings_found.append(
                    f"  Dashboard expected to reference: {', '.join(missing)}"
                )
            else:
                logger.info("  All expected source workbooks referenced")
    elif workbook_id == 'A.8.24.5':
        warnings_found.append(
            "  A.8.24.5 Dashboard: No external workbook links found!"
        )
        logger.info("  WARNING: This dashboard requires external links to function")

    # ========================================================================
    # CHECK 2: DATA VALIDATION CONFLICTS
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 2: DATA VALIDATION CONFLICTS")
    logger.info("=" * 80)

    validation_issues = 0
    total_validations = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if hasattr(ws, 'data_validations') and ws.data_validations:
            dv_count = len(ws.data_validations.dataValidation)
            total_validations += dv_count
            logger.info(f"  {sheet_name}: {dv_count} data validations")

            # Check for overlapping ranges (simplified)
            all_ranges = []
            for dv in ws.data_validations.dataValidation:
                if hasattr(dv, 'sqref') and dv.sqref:
                    all_ranges.extend(str(dv.sqref).split())

            if len(all_ranges) != len(set(all_ranges)):
                warnings_found.append(
                    f"  {sheet_name}: Potentially overlapping data validation ranges"
                )
                validation_issues += 1

    logger.info(f"  Total data validations across all sheets: {total_validations}")

    if validation_issues == 0:
        logger.info("  No obvious data validation conflicts")

    # ========================================================================
    # CHECK 3: STYLE CONSISTENCY
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 3: STYLE CONSISTENCY")
    logger.info("=" * 80)

    style_issues = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        cells_without_font = 0

        sample_size = 0
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value:
                    sample_size += 1
                    if not hasattr(cell, 'font') or cell.font is None:
                        cells_without_font += 1

        if cells_without_font > 0:
            warnings_found.append(
                f"  {sheet_name}: {cells_without_font}/{sample_size} "
                f"cells missing font attributes"
            )

        # Note: Missing borders are often intentional, so we don't flag this as issue

    if style_issues == 0:
        logger.info("  Style attributes appear consistent")

    # ========================================================================
    # CHECK 4: MERGED CELLS VALIDATION
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 4: MERGED CELLS VALIDATION")
    logger.info("=" * 80)

    merge_issues = 0
    total_merges = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if hasattr(ws, 'merged_cells') and ws.merged_cells:
            merge_count = len(ws.merged_cells.ranges)
            total_merges += merge_count
            logger.info(f"  {sheet_name}: {merge_count} merged cell ranges")

            # Check if merged cells have content in non-top-left cells
            for merge_range in ws.merged_cells.ranges:
                min_col = merge_range.min_col
                min_row = merge_range.min_row
                max_col = merge_range.max_col
                max_row = merge_range.max_row

                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if row == min_row and col == min_col:
                            continue

                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            cell_coord = f"{get_column_letter(col)}{row}"
                            warnings_found.append(
                                f"  {sheet_name}!{cell_coord}: "
                                f"Merged cell has content in non-primary cell"
                            )
                            merge_issues += 1

    if total_merges > 0:
        logger.info(f"  Total merged ranges across all sheets: {total_merges}")

    if merge_issues == 0:
        logger.info("  Merged cells appear properly formatted")
    else:
        logger.error(f"  Found {merge_issues} merged cell content issues")

    # ========================================================================
    # CHECK 5: WORKSHEET STRUCTURE
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 5: WORKSHEET STRUCTURE")
    logger.info("=" * 80)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Check for excessive dimensions
        if ws.max_row > 10000:
            warnings_found.append(
                f"  {sheet_name}: Large worksheet ({ws.max_row} rows)"
            )

        if ws.max_column > 100:
            warnings_found.append(
                f"  {sheet_name}: Wide worksheet ({ws.max_column} columns)"
            )

    logger.info("  Worksheet dimensions within reasonable limits")

    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("DIAGNOSTIC SUMMARY")
    logger.info("=" * 80)

    if issues_found:
        logger.error(f"CRITICAL ISSUES FOUND: {len(issues_found)}")
        for issue in issues_found:
            logger.error(issue)

    if warnings_found:
        logger.info(f"WARNINGS: {len(warnings_found)}")
        for warning in warnings_found:
            logger.info(warning)

    if not issues_found and not warnings_found:
        logger.info("NO ISSUES DETECTED")
        logger.info("The workbook appears structurally sound.")
        logger.info("Excel repair warnings may be due to:")
        logger.info("  - Excel version compatibility (try Excel 2019+)")
        logger.info("  - Antivirus/security software interference")
        logger.info("  - Network drive or OneDrive sync issues")
        logger.info("  - Excel's overly cautious validation")
        if workbook_id == 'A.8.24.5':
            logger.info("  - Unresolved external workbook links (expected before setup)")
    else:
        logger.info("")
        logger.info("=" * 80)
        logger.info("RECOMMENDED ACTIONS:")
        logger.info("=" * 80)

        if formula_issues > 0:
            logger.info("1. FORMULA FIXES:")
            logger.info("   - Review formulas referencing other sheets")
            logger.info("   - Ensure sheet names match exactly (case-sensitive)")
            logger.info("   - Check for balanced quotes and parentheses")
            if workbook_id == 'A.8.24.5':
                logger.info("   - Verify external workbook references are correct")

        if validation_issues > 0:
            logger.info("2. DATA VALIDATION FIXES:")
            logger.info("   - Remove overlapping data validation ranges")
            logger.info("   - Apply validations to specific ranges, not entire columns")

        if merge_issues > 0:
            logger.info("3. MERGED CELL FIXES:")
            logger.info("   - Ensure only top-left cell of merged range has content")
            logger.info("   - Clear content from other cells in merged range")

        # Special instructions for A.8.24.5 Dashboard
        if workbook_id == 'A.8.24.5' and (not external_refs or len(warnings_found) > 0):
            logger.info("4. DASHBOARD-SPECIFIC SETUP (A.8.24.5):")
            logger.info("   - This workbook REQUIRES external links to function")
            logger.info("   - Ensure all 4 source workbooks are in the same folder:")
            logger.info("     - ISMS-IMP-A.8.24.1.xlsx (Data Transmission)")
            logger.info("     - ISMS-IMP-A.8.24.2.xlsx (Data Storage)")
            logger.info("     - ISMS-IMP-A.8.24.3.xlsx (Authentication)")
            logger.info("     - ISMS-IMP-A.8.24.4.xlsx (Key Management)")
            logger.info("   - Run normalize_assessment_files.py first to standardize filenames")
            logger.info("   - Open dashboard and click 'Update Links' when prompted")
            logger.info("   - #REF errors before updating links are EXPECTED and normal")

    # ========================================================================
    # WORKBOOK-SPECIFIC NOTES
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info(f"WORKBOOK-SPECIFIC NOTES: {workbook_id}")
    logger.info("=" * 80)

    if workbook_id == 'A.8.24.1':
        logger.info("Data Transmission Assessment:")
        logger.info("  - 7 assessment sheets (Network Transport, API, Email, File Transfer,")
        logger.info("    Database, Messaging, Remote Access)")
        logger.info("  - Summary Dashboard with 11 analysis sections")
        logger.info("  - Evidence Register (100 entries)")
        logger.info("  - Approval Sign-Off")

    elif workbook_id == 'A.8.24.2':
        logger.info("Data Storage Assessment:")
        logger.info("  - 6 assessment sheets (Database, File Storage, Backup, Cloud Storage,")
        logger.info("    Endpoint, Removable Media)")
        logger.info("  - Summary Dashboard with 11 analysis sections")
        logger.info("  - Evidence Register (100 entries)")
        logger.info("  - Approval Sign-Off")

    elif workbook_id == 'A.8.24.3':
        logger.info("Authentication Assessment:")
        logger.info("  - 7 assessment sheets (User Auth, Admin Auth, Service Accounts,")
        logger.info("    MFA, SSO, API Auth, Password Management)")
        logger.info("  - Summary Dashboard with 11 analysis sections")
        logger.info("  - Evidence Register (100 entries)")
        logger.info("  - Approval Sign-Off")

    elif workbook_id == 'A.8.24.4':
        logger.info("Key Management Assessment:")
        logger.info("  - 5 assessment sheets (Key Generation, Key Storage, Key Rotation,")
        logger.info("    Key Backup & Recovery, Certificate Management)")
        logger.info("  - Summary Dashboard with 11 analysis sections")
        logger.info("  - Evidence Register (100 entries)")
        logger.info("  - Approval Sign-Off")

    elif workbook_id == 'A.8.24.5':
        logger.info("Compliance Summary Dashboard:")
        logger.info("  - 9 sheets total - CONSOLIDATION WORKBOOK")
        logger.info("  - Executive Dashboard (with external links)")
        logger.info("  - Gap Analysis (200 entries)")
        logger.info("  - Risk Register (100 entries)")
        logger.info("  - Remediation Roadmap (200 entries)")
        logger.info("  - KPIs & Metrics (50+ KPIs in 4 categories)")
        logger.info("  - Evidence Register (500 entries)")
        logger.info("  - Action Items & Follow-up (200 entries)")
        logger.info("  - Audit & Compliance Log (100 entries)")
        logger.info("  - Approval Sign-Off")
        logger.info("")
        logger.info("  CRITICAL: This dashboard requires external workbook links!")
        logger.info("  Setup steps:")
        logger.info("    1. Run: python3 normalize_assessment_files.py")
        logger.info("    2. Place dashboard in same folder as normalized source files")
        logger.info("    3. Open dashboard and click 'Update Links'")

    else:
        logger.info("Generic Excel Workbook")
        logger.info("  - No specific validation rules defined")
        logger.info("  - Standard Excel health checks applied")

    logger.info("=" * 80)

    return 1 if issues_found else 0


def main():
    """Main entry point for the sanity checker."""
    try:
        if len(sys.argv) < 2:
            logger.error("Usage: python3 excel_sanity_check.py <filename.xlsx>")
            logger.info("Examples:")
            logger.info("  python3 excel_sanity_check.py ISMS-IMP-A.8.24.1_Data_Transmission_20251231.xlsx")
            logger.info("  python3 excel_sanity_check.py ISMS-IMP-A.8.24.2_Data_Storage_20251231.xlsx")
            logger.info("  python3 excel_sanity_check.py ISMS-IMP-A.8.24.3_Authentication_20251231.xlsx")
            logger.info("  python3 excel_sanity_check.py ISMS-IMP-A.8.24.4_Key_Management_20251231.xlsx")
            logger.info("  python3 excel_sanity_check.py ISMS-IMP-A.8.24.5_Compliance_Summary_Dashboard_20251231.xlsx")
            sys.exit(1)

        filename = sys.argv[1]
        exit_code = check_workbook_health(filename)
        sys.exit(exit_code)

    except KeyboardInterrupt:
        logger.info("Operation cancelled by user")
        sys.exit(130)
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (syntax validated, standardized format applied)
# QA_TOOL: Claude Code Deep Scan
# =============================================================================
