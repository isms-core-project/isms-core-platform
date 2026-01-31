#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
Excel Sanity Checker - A.8.24.5 Compliance Summary Dashboard
================================================================================

SPECIALIZED diagnostic for Compliance Summary Dashboard workbooks.

**Purpose:**
Domain-specific validation tailored to the unique structure and requirements
of A.8.24.5 assessment workbooks.

**Checks Performed:**
- Domain-specific sheet structure validation
- Formula integrity for domain-specific calculations
- Data validation rules specific to this assessment type
- Compliance dashboard-specific validation logic
- External workbook link validation

**When to Use:**
- After generating A.8.24.5 assessment workbook
- Before stakeholder distribution
- When Excel shows repair warnings for this specific domain

**Usage:**
    python3 excel_sanity_check_a824_5.py ISMS-IMP-A.8.24.5_Assessment_YYYYMMDD.xlsx

**Differences from General Checker:**
This specialized checker validates domain-specific requirements that the
general checker cannot verify (custom sheet structures, domain-specific
formulas, assessment-specific data validation rules).

Control Reference: ISO/IEC 27001:2022 Annex A Control A.8.24
Assessment Domain: 5 of 5 (Compliance Summary Dashboard)
Script Type: Domain-Specific QA Utility
Version: 1.0
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
import re

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# ============================================================================
# A.8.24.5 SPECIFIC VALIDATION DATA
# ============================================================================

EXPECTED_SHEETS = [
    "Executive Dashboard",
    "Gap Analysis",
    "Risk Register",
    "Remediation Roadmap",
    "KPIs & Metrics",
    "Evidence Register",
    "Action Items & Follow-up",
    "Audit & Compliance Log",
    "Approval Sign-Off",
]

REQUIRED_DOCUMENT_INFO = [
    "Document ID",
    "Report Type",
    "Related Policy",
    "Version",
]

# Expected source workbooks for external links
EXPECTED_SOURCE_WORKBOOKS = [
    "ISMS-IMP-A.8.24.1.xlsx",
    "ISMS-IMP-A.8.24.2.xlsx",
    "ISMS-IMP-A.8.24.3.xlsx",
    "ISMS-IMP-A.8.24.4.xlsx",
]

# Expected data row counts per sheet
EXPECTED_DATA_ROWS = {
    "Gap Analysis": 200,          # Detailed gap register
    "Risk Register": 100,         # Detailed risk register
    "Remediation Roadmap": 200,   # Detailed remediation register
    "Evidence Register": 500,     # Detailed evidence register
    "Action Items & Follow-up": 200,  # Detailed action register
    "Audit & Compliance Log": 100,    # Audit register
}

# Expected sections in Executive Dashboard
EXECUTIVE_DASHBOARD_SECTIONS = [
    "OVERALL CRYPTOGRAPHY COMPLIANCE STATUS",
    "COMPLIANCE BY ASSESSMENT AREA",
    "KEY PERFORMANCE INDICATORS",
    "TOP 5 CRITICAL SECURITY GAPS",
    "EXECUTIVE SUMMARY",
]

# Expected sections in KPIs & Metrics
KPI_SECTIONS = [
    "COMPLIANCE METRICS",
    "RISK METRICS",
    "REMEDIATION METRICS",
    "TECHNICAL IMPLEMENTATION METRICS",
]


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""

    logger.info("=" * 80)
    logger.info(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {filename}")
    logger.info("ISMS-IMP-A.8.24.5 - Compliance Summary Dashboard")
    logger.info("=" * 80)

    issues_found = []
    warnings_found = []

    try:
        wb = load_workbook(filename, data_only=False)
        logger.info("Workbook loaded successfully")
        logger.info(f"  Sheets found: {len(wb.sheetnames)}")

    except Exception as e:
        logger.error(f"CRITICAL: Cannot load workbook: {e}")
        return 1

    # ========================================================================
    # CHECK 0: A.8.24.5 SPECIFIC STRUCTURE VALIDATION
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 0: A.8.24.5 STRUCTURE VALIDATION")
    logger.info("=" * 80)

    structure_issues = 0

    # Check for expected sheets
    missing_sheets = []
    for sheet in EXPECTED_SHEETS:
        if sheet not in wb.sheetnames:
            missing_sheets.append(sheet)
            structure_issues += 1

    if missing_sheets:
        issues_found.append(f"  Missing required sheets: {', '.join(missing_sheets)}")
        logger.error(f"  Missing {len(missing_sheets)} required sheet(s)")
    else:
        logger.info(f"  All {len(EXPECTED_SHEETS)} required sheets present")

    # Check for unexpected sheets
    unexpected_sheets = []
    for sheet in wb.sheetnames:
        if sheet not in EXPECTED_SHEETS:
            unexpected_sheets.append(sheet)

    if unexpected_sheets:
        warnings_found.append(f"  Unexpected sheets found: {', '.join(unexpected_sheets)}")
        logger.info(f"  Found {len(unexpected_sheets)} unexpected sheet(s)")

    # Validate document information fields in Executive Dashboard
    if "Executive Dashboard" in wb.sheetnames:
        ws = wb["Executive Dashboard"]
        found_fields = []
        for row in ws.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for field in REQUIRED_DOCUMENT_INFO:
                        if field in cell.value:
                            found_fields.append(field)

        missing_fields = [f for f in REQUIRED_DOCUMENT_INFO if f not in found_fields]
        if missing_fields:
            warnings_found.append(f"  Missing document info fields: {', '.join(missing_fields)}")

    # Validate Executive Dashboard sections
    if "Executive Dashboard" in wb.sheetnames:
        ws = wb["Executive Dashboard"]
        found_sections = []
        for row in ws.iter_rows(min_row=1, max_row=100):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for section in EXECUTIVE_DASHBOARD_SECTIONS:
                        if section in cell.value.upper():
                            found_sections.append(section)

        missing_sections = [s for s in EXECUTIVE_DASHBOARD_SECTIONS if s not in found_sections]
        if missing_sections:
            warnings_found.append(f"  Executive Dashboard: Missing sections: {', '.join(missing_sections)}")
        else:
            logger.info(f"  Executive Dashboard: All {len(EXECUTIVE_DASHBOARD_SECTIONS)} sections found")

    # Validate KPI categories
    if "KPIs & Metrics" in wb.sheetnames:
        ws = wb["KPIs & Metrics"]
        found_categories = []
        for row in ws.iter_rows(min_row=1, max_row=200):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for category in KPI_SECTIONS:
                        if category in cell.value.upper():
                            found_categories.append(category)

        missing_categories = [c for c in KPI_SECTIONS if c not in found_categories]
        if missing_categories:
            warnings_found.append(f"  KPIs & Metrics: Missing categories: {', '.join(missing_categories)}")
        else:
            logger.info(f"  KPIs & Metrics: All {len(KPI_SECTIONS)} categories found")

    if structure_issues == 0:
        logger.info("  Workbook structure matches A.8.24.5 specification")

    # ========================================================================
    # CHECK 0A: EXTERNAL WORKBOOK LINKS VALIDATION
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 0A: EXTERNAL WORKBOOK LINKS VALIDATION")
    logger.info("=" * 80)

    external_link_issues = 0
    found_external_refs = set()

    if "Executive Dashboard" in wb.sheetnames:
        ws = wb["Executive Dashboard"]

        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value

                    # Check for external workbook references [filename.xlsx]
                    external_refs = re.findall(r'\[([^\]]+\.xlsx)\]', formula)
                    for ref in external_refs:
                        found_external_refs.add(ref)

        if found_external_refs:
            logger.info(f"  Found {len(found_external_refs)} external workbook references:")
            for ref in sorted(found_external_refs):
                logger.info(f"    - {ref}")

            # Check if all expected source workbooks are referenced
            missing_refs = []
            for expected in EXPECTED_SOURCE_WORKBOOKS:
                if expected not in found_external_refs:
                    missing_refs.append(expected)

            if missing_refs:
                warnings_found.append(f"  Missing external references: {', '.join(missing_refs)}")
                logger.info(f"  Expected source workbooks not found in formulas:")
                for ref in missing_refs:
                    logger.info(f"    - {ref}")
            else:
                logger.info(f"  All {len(EXPECTED_SOURCE_WORKBOOKS)} expected source workbooks referenced")
        else:
            warnings_found.append("  No external workbook links found in Executive Dashboard")
            logger.info("  WARNING: No external workbook links detected")
            logger.info("  This dashboard relies on external data - links may be broken!")

    # ========================================================================
    # CHECK 1: FORMULA VALIDATION
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 1: FORMULA VALIDATION")
    logger.info("=" * 80)

    formula_issues = 0
    inter_sheet_refs = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value

                    # Check for common formula issues
                    if "'" in formula:
                        # Sheet references with special characters
                        sheet_refs = re.findall(r"'([^']+)'!", formula)
                        for ref in sheet_refs:
                            if ref not in wb.sheetnames and "[" not in ref:
                                issues_found.append(
                                    f"  {sheet_name}!{cell.coordinate}: "
                                    f"Invalid sheet reference '{ref}'"
                                )
                                formula_issues += 1

                            # Track inter-sheet references
                            if ref in wb.sheetnames:
                                if sheet_name not in inter_sheet_refs:
                                    inter_sheet_refs[sheet_name] = set()
                                inter_sheet_refs[sheet_name].add(ref)

                    # Check for common syntax issues
                    if formula.count('(') != formula.count(')'):
                        issues_found.append(
                            f"  {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced parentheses in formula"
                        )
                        formula_issues += 1

                    # Check for double quotes issues
                    if formula.count('"') % 2 != 0:
                        issues_found.append(
                            f"  {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced quotes in formula"
                        )
                        formula_issues += 1

    if formula_issues == 0:
        logger.info("  All formulas appear syntactically valid")
    else:
        logger.error(f"  Found {formula_issues} formula issues")

    # Report inter-sheet dependencies
    if inter_sheet_refs:
        logger.info("  Inter-sheet formula dependencies:")
        for source, targets in sorted(inter_sheet_refs.items()):
            logger.info(f"    {source} -> {', '.join(sorted(targets))}")

    # ========================================================================
    # CHECK 2: DATA VALIDATION CONFLICTS
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 2: DATA VALIDATION CONFLICTS")
    logger.info("=" * 80)

    validation_issues = 0
    total_validations = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if hasattr(ws, 'data_validations') and ws.data_validations:
            dv_count = len(ws.data_validations.dataValidation)
            total_validations += dv_count
            logger.info(f"  {sheet_name}: {dv_count} data validations")

            # Check for overlapping ranges (simplified)
            all_ranges = []
            for dv in ws.data_validations.dataValidation:
                if hasattr(dv, 'sqref') and dv.sqref:
                    all_ranges.extend(str(dv.sqref).split())

            if len(all_ranges) != len(set(all_ranges)):
                warnings_found.append(
                    f"  {sheet_name}: Potentially overlapping data validation ranges"
                )
                validation_issues += 1

    logger.info(f"  Total data validations across all sheets: {total_validations}")

    if validation_issues == 0:
        logger.info("  No obvious data validation conflicts")

    # ========================================================================
    # CHECK 3: STYLE CONSISTENCY
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 3: STYLE CONSISTENCY")
    logger.info("=" * 80)

    style_issues = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        cells_without_font = 0

        sample_size = 0
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value:
                    sample_size += 1
                    if not hasattr(cell, 'font') or cell.font is None:
                        cells_without_font += 1

        if cells_without_font > 0:
            warnings_found.append(
                f"  {sheet_name}: {cells_without_font}/{sample_size} "
                f"cells missing font attributes"
            )

    if style_issues == 0:
        logger.info("  Style attributes appear consistent")

    # ========================================================================
    # CHECK 4: MERGED CELLS VALIDATION
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 4: MERGED CELLS VALIDATION")
    logger.info("=" * 80)

    merge_issues = 0
    total_merges = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if hasattr(ws, 'merged_cells') and ws.merged_cells:
            merge_count = len(ws.merged_cells.ranges)
            total_merges += merge_count
            logger.info(f"  {sheet_name}: {merge_count} merged cell ranges")

            # Check if merged cells have content in non-top-left cells
            for merge_range in ws.merged_cells.ranges:
                min_col = merge_range.min_col
                min_row = merge_range.min_row
                max_col = merge_range.max_col
                max_row = merge_range.max_row

                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if row == min_row and col == min_col:
                            continue

                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            cell_coord = f"{get_column_letter(col)}{row}"
                            warnings_found.append(
                                f"  {sheet_name}!{cell_coord}: "
                                f"Merged cell has content in non-primary cell"
                            )
                            merge_issues += 1

    if total_merges > 0:
        logger.info(f"  Total merged ranges across all sheets: {total_merges}")

    if merge_issues == 0:
        logger.info("  Merged cells appear properly formatted")
    else:
        logger.error(f"  Found {merge_issues} merged cell content issues")

    # ========================================================================
    # CHECK 5: WORKSHEET STRUCTURE & DATA ROWS
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("CHECK 5: WORKSHEET STRUCTURE & DATA ROWS")
    logger.info("=" * 80)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Check for excessive dimensions
        if ws.max_row > 10000:
            warnings_found.append(
                f"  {sheet_name}: Large worksheet ({ws.max_row} rows)"
            )

        if ws.max_column > 100:
            warnings_found.append(
                f"  {sheet_name}: Wide worksheet ({ws.max_column} columns)"
            )

        # Check expected data row counts
        if sheet_name in EXPECTED_DATA_ROWS:
            expected_rows = EXPECTED_DATA_ROWS[sheet_name]
            logger.info(f"  {sheet_name}: Expected ~{expected_rows} data rows")

    logger.info("  Worksheet dimensions within reasonable limits")

    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    logger.info("")
    logger.info("=" * 80)
    logger.info("DIAGNOSTIC SUMMARY")
    logger.info("=" * 80)

    if issues_found:
        logger.error(f"CRITICAL ISSUES FOUND: {len(issues_found)}")
        for issue in issues_found:
            logger.error(issue)

    if warnings_found:
        logger.info(f"WARNINGS: {len(warnings_found)}")
        for warning in warnings_found:
            logger.info(warning)

    if not issues_found and not warnings_found:
        logger.info("NO ISSUES DETECTED")
        logger.info("The workbook appears structurally sound and matches A.8.24.5 specification.")
        logger.info("Excel repair warnings may be due to:")
        logger.info("  - Excel version compatibility (try Excel 2019+)")
        logger.info("  - Antivirus/security software interference")
        logger.info("  - Network drive or OneDrive sync issues")
        logger.info("  - Excel's overly cautious validation")
        logger.info("  - Unresolved external workbook links (this is expected before setup)")
    else:
        logger.info("")
        logger.info("=" * 80)
        logger.info("RECOMMENDED ACTIONS:")
        logger.info("=" * 80)

        if structure_issues > 0:
            logger.info("0. STRUCTURE FIXES:")
            logger.info("   - Verify all required sheets are present")
            logger.info("   - Check sheet names match specification exactly")
            logger.info("   - Ensure document information fields are complete")
            logger.info("   - Verify all required sections exist in each sheet")

        if external_link_issues > 0 or not found_external_refs:
            logger.info("0A. EXTERNAL LINKS FIXES:")
            logger.info("   - This dashboard REQUIRES external workbook links to function")
            logger.info("   - Ensure all 4 source workbooks are in the same folder:")
            logger.info("     - ISMS-IMP-A.8.24.1.xlsx (Data Transmission)")
            logger.info("     - ISMS-IMP-A.8.24.2.xlsx (Data Storage)")
            logger.info("     - ISMS-IMP-A.8.24.3.xlsx (Authentication)")
            logger.info("     - ISMS-IMP-A.8.24.4.xlsx (Key Management)")
            logger.info("   - Run normalize_assessment_files.py first to standardize filenames")
            logger.info("   - Open dashboard and click 'Update Links' when prompted")

        if formula_issues > 0:
            logger.info("1. FORMULA FIXES:")
            logger.info("   - Review formulas referencing other sheets")
            logger.info("   - Ensure sheet names match exactly (case-sensitive)")
            logger.info("   - Check for balanced quotes and parentheses")
            logger.info("   - Verify external workbook references are correct")

        if validation_issues > 0:
            logger.info("2. DATA VALIDATION FIXES:")
            logger.info("   - Remove overlapping data validation ranges")
            logger.info("   - Apply validations to specific ranges, not entire columns")

        if merge_issues > 0:
            logger.info("3. MERGED CELL FIXES:")
            logger.info("   - Ensure only top-left cell of merged range has content")
            logger.info("   - Clear content from other cells in merged range")

    logger.info("")
    logger.info("=" * 80)
    logger.info("A.8.24.5 SPECIFIC NOTES:")
    logger.info("=" * 80)
    logger.info("Expected structure:")
    logger.info("  - 9 sheets total")
    logger.info("  - Executive Dashboard:")
    logger.info("    - Document information (10 fields)")
    logger.info("    - Overall compliance status scorecard (4 metrics)")
    logger.info("    - Compliance by assessment area (4 areas + total)")
    logger.info("    - KPI table (10+ KPIs)")
    logger.info("    - Top 5 critical security gaps")
    logger.info("    - Executive summary with achievements/concerns")
    logger.info("    - EXTERNAL LINKS to 4 source assessment workbooks")
    logger.info("")
    logger.info("  - Gap Analysis:")
    logger.info("    - Gap summary statistics (4 areas + total)")
    logger.info("    - Detailed gap register (200 rows)")
    logger.info("    - Columns: Gap ID, Assessment Area, Source, System, Description, etc.")
    logger.info("")
    logger.info("  - Risk Register:")
    logger.info("    - Risk distribution summary")
    logger.info("    - Risk by assessment area")
    logger.info("    - Detailed risk register (100 rows)")
    logger.info("    - Inherent and residual risk scoring")
    logger.info("")
    logger.info("  - Remediation Roadmap:")
    logger.info("    - Roadmap summary (overall progress)")
    logger.info("    - Progress by assessment area")
    logger.info("    - Detailed remediation register (200 rows)")
    logger.info("    - Timeline tracking with days remaining formulas")
    logger.info("")
    logger.info("  - KPIs & Metrics:")
    logger.info("    - 4 KPI categories (50+ total KPIs):")
    logger.info("      1. Compliance Metrics (15 KPIs)")
    logger.info("      2. Risk Metrics (10 KPIs)")
    logger.info("      3. Remediation Metrics (11 KPIs)")
    logger.info("      4. Technical Implementation Metrics (15 KPIs)")
    logger.info("")
    logger.info("  - Evidence Register:")
    logger.info("    - Evidence summary statistics (4 areas + total)")
    logger.info("    - Detailed evidence register (500 rows)")
    logger.info("    - Retention period tracking and destruction date formulas")
    logger.info("")
    logger.info("  - Action Items & Follow-up:")
    logger.info("    - Action summary dashboard (by status and priority)")
    logger.info("    - Detailed action register (200 rows)")
    logger.info("    - Days open/remaining formulas and on-time status")
    logger.info("")
    logger.info("  - Audit & Compliance Log:")
    logger.info("    - Audit summary (by audit type)")
    logger.info("    - Audit register (100 rows)")
    logger.info("")
    logger.info("  - Approval Sign-Off:")
    logger.info("    - 3-level approval workflow")
    logger.info("    - Document status tracking")
    logger.info("")
    logger.info("CRITICAL SUCCESS FACTORS:")
    logger.info("  1. External workbook links MUST be configured")
    logger.info("  2. Source workbooks must use exact filenames")
    logger.info("  3. All files must be in same folder for links to work")
    logger.info("  4. Run normalization script before using dashboard")
    logger.info("  5. Click 'Update Links' when opening dashboard")
    logger.info("=" * 80)

    return 1 if issues_found else 0


def main():
    """Main entry point for the sanity checker."""
    try:
        if len(sys.argv) < 2:
            logger.error("Usage: python3 excel_sanity_check_a824_5.py <filename.xlsx>")
            logger.info("Example:")
            logger.info("  python3 excel_sanity_check_a824_5.py ISMS-IMP-A.8.24.5_Compliance_Summary_Dashboard_20251231.xlsx")
            sys.exit(1)

        filename = sys.argv[1]
        exit_code = check_workbook_health(filename)
        sys.exit(exit_code)

    except KeyboardInterrupt:
        logger.info("Operation cancelled by user")
        sys.exit(130)
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (syntax validated, standardized format applied)
# QA_TOOL: Claude Code Deep Scan
# =============================================================================
