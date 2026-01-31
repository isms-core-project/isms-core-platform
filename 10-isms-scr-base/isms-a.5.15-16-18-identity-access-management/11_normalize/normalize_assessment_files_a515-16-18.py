#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
normalize_assessment_files_a515-16-18.py
================================================================================

ISMS Assessment File Normalization Utility
Prepares A.5.15-16-18 IAM assessment workbooks for dashboard integration

--------------------------------------------------------------------------------
UTILITY SCRIPT - MINIMAL CUSTOMIZATION REQUIRED
--------------------------------------------------------------------------------

This script is a SUPPORT UTILITY with minimal customization needs. It normalizes
assessment workbook filenames to enable dashboard external workbook linking.

Key customization areas:
1. EXPECTED_DOCS dictionary (if you rename assessment files)
2. File path locations (if not using current directory)
3. Output directory for normalized files (optional)

Most users can run this script WITHOUT customization.

Reference Pattern: Based on ISMS normalization utility pattern

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script normalizes A.5.15-16-18 assessment workbook filenames by creating
copies with standardized names (no dates or version suffixes), enabling the
dashboard to maintain stable external workbook references across assessment updates.

**Purpose:**
Solves the "Excel #REF! error problem" when dashboard links break every time
assessment files are regenerated with new dates. Creates normalized copies with
fixed filenames for reliable dashboard external workbook linking.

**The Problem This Solves:**

Without normalization:
```
1. Generate assessment: ISMS-IMP-A.5.15-16-18.S1_20260110.xlsx
2. Dashboard links to:  ISMS-IMP-A.5.15-16-18.S1_20260110.xlsx
3. Update assessment:   ISMS-IMP-A.5.15-16-18.S1_20260115.xlsx (new date!)
4. Dashboard breaks:    #REF! errors (looking for *_20260110.xlsx)
```

With normalization:
```
1. Generate assessment:  ISMS-IMP-A.5.15-16-18.S1_20260110.xlsx
2. Run normalization:    Creates ISMS-IMP-A.5.15-16-18.S1.xlsx (no date)
3. Dashboard links to:   ISMS-IMP-A.5.15-16-18.S1.xlsx (stable name)
4. Update assessment:    ISMS-IMP-A.5.15-16-18.S1_20260115.xlsx
5. Re-run normalization: Updates ISMS-IMP-A.5.15-16-18.S1.xlsx
6. Dashboard works:      No broken links!
```

**Key Principle:** Dashboard always reads from normalized files (no dates).
Original dated files preserved for version history and audit trail.

**Process Flow:**
1. Scan directory for assessment workbooks with date suffixes
2. Validate each workbook structure (Instructions & Legend sheet present)
3. Identify most recent version of each assessment type
4. Copy to normalized filename (without date suffix)
5. Create normalization manifest (audit trail of what was normalized)
6. Generate console report with file locations

**Normalized Filenames Created:**
- `ISMS-IMP-A.5.15-16-18.S1.xlsx` (User Inventory & Lifecycle Compliance)
- `ISMS-IMP-A.5.15-16-18.S2.xlsx` (Access Rights Matrix)
- `ISMS-IMP-A.5.15-16-18.S3.xlsx` (Access Review Results)
- `ISMS-IMP-A.5.15-16-18.S4.xlsx` (Role Compliance & SoD)
- `ISMS-IMP-A.5.15-16-18.S5.xlsx` (Lifecycle Compliance Detailed)

**Output Structure:**
```
/assessment-directory/
├── ISMS-IMP-A.5.15-16-18.S1_20260110.xlsx    (original - preserved)
├── ISMS-IMP-A.5.15-16-18.S1_20260115.xlsx    (original - preserved)
├── ISMS-IMP-A.5.15-16-18.S1.xlsx             (normalized - dashboard reads this)
├── ISMS-IMP-A.5.15-16-18.S2_20260110.xlsx    (original - preserved)
├── ISMS-IMP-A.5.15-16-18.S2.xlsx             (normalized - dashboard reads this)
├── ISMS-IMP-A.5.15-16-18.S3_20260110.xlsx    (original - preserved)
├── ISMS-IMP-A.5.15-16-18.S3.xlsx             (normalized - dashboard reads this)
├── ISMS-IMP-A.5.15-16-18.S4_20260110.xlsx    (original - preserved)
├── ISMS-IMP-A.5.15-16-18.S4.xlsx             (normalized - dashboard reads this)
├── ISMS-IMP-A.5.15-16-18.S5_20260110.xlsx    (original - preserved)
├── ISMS-IMP-A.5.15-16-18.S5.xlsx             (normalized - dashboard reads this)
└── Normalization_Manifest_20260115_143022.txt (audit trail)
```

**Key Features:**
- Automatic detection of latest assessment versions by date
- Workbook structure validation before normalization
- Preserved original files (non-destructive operation)
- Audit manifest generation (traceability)
- Comprehensive console reporting (which files normalized)
- Error handling for missing or corrupt workbooks
- UTF-8 encoding support for emoji status indicators
- Flexible directory operation (current or specified path)

**Integration:**
This utility is REQUIRED before running IAM Governance Dashboard generator
to ensure dashboard external workbook links resolve correctly. Without
normalization, dashboard will show #REF! errors.

Workflow:
1. Complete all 5 assessment workbooks (fill in data)
2. Run: `python3 normalize_assessment_files_a515-16-18.py`
3. Run: `python3 generate_a515-16-18_dashboard.py`
4. Dashboard successfully reads normalized files

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library (for workbook validation)

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl --break-system-packages

Dependencies:
    - openpyxl >= 3.1.0 (Python Excel library - for validation)
    - datetime (standard library)
    - os (standard library)
    - sys (standard library)
    - shutil (standard library - for file copying)
    - pathlib (standard library - for path handling)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage (Scan Current Directory):
    python3 normalize_assessment_files_a515-16-18.py

Specify Directory:
    python3 normalize_assessment_files_a515-16-18.py /path/to/assessments

Output Location:
    Normalized files created in SAME directory as source files
    (keeps dashboard and source files together)

Typical Workflow:
```bash
# Step 1: Complete assessments
python3 generate_a515-16-18_1_user_inventory.py
python3 generate_a515-16-18_2_access_rights_matrix.py
python3 generate_a515-16-18_3_access_review_results.py
python3 generate_a515-16-18_4_role_compliance.py
python3 generate_a515-16-18_5_lifecycle_compliance.py

# Step 2: Fill out assessment workbooks (manual step)
# [Users document identities, access rights, reviews, etc.]

# Step 3: Normalize filenames
python3 normalize_assessment_files_a515-16-18.py

# Step 4: Generate dashboard
python3 generate_a515-16-18_dashboard.py

# Step 5: Dashboard successfully links to normalized files
# No #REF! errors!
```

When to Re-Run:
- After regenerating any assessment workbook with new date
- After updating assessment content (to update normalized copy)
- Before generating/regenerating dashboard
- After organizational assessment cycles

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.5.15, A.5.16, A.5.18
Utility Type:         Quality Assurance - Filename Normalization
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.15-16-18: IAM Policy Framework (Governance)
    - ISMS-IMP-A.5.15-16-18: IAM Implementation Guides

Related Assessments:
    - ISMS-IMP-A.5.15-16-18.S1: User Inventory Assessment
    - ISMS-IMP-A.5.15-16-18.S2: Access Rights Matrix Assessment
    - ISMS-IMP-A.5.15-16-18.S3: Access Review Results Assessment
    - ISMS-IMP-A.5.15-16-18.S4: Role Compliance Assessment
    - ISMS-IMP-A.5.15-16-18.S5: Lifecycle Compliance Detailed Assessment
    - IAM Governance Dashboard (REQUIRES normalized files)

Related Scripts:
    - generate_a515-16-18_1_user_inventory.py (Assessment 1 generator)
    - generate_a515-16-18_2_access_rights_matrix.py (Assessment 2 generator)
    - generate_a515-16-18_3_access_review_results.py (Assessment 3 generator)
    - generate_a515-16-18_4_role_compliance.py (Assessment 4 generator)
    - generate_a515-16-18_5_lifecycle_compliance.py (Assessment 5 generator)
    - generate_a515-16-18_dashboard.py (Dashboard - REQUIRES normalized files)
    - excel_sanity_check_a515.py (Workbook validation utility)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements filename normalization for dashboard external link stability
    - Creates audit manifest for traceability
    - Validates workbook structure before normalization
    - Supports flexible directory operation

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Why Normalization is Critical:**

Excel external workbook links are FRAGILE. They break when source files are
renamed, moved, or dated. Dashboard consolidation REQUIRES stable filenames.

Without this utility:
- Every assessment regeneration breaks dashboard links
- #REF! errors appear in dashboard formulas
- Manual relinking required (error-prone, time-consuming)
- Dashboard becomes maintenance nightmare

With this utility:
- Dashboard always reads from stable normalized filenames
- Assessment updates don't break links
- Re-run normalization after updates → dashboard automatically picks up changes
- Zero manual relinking required

**Non-Destructive Operation:**

This script COPIES files, it does NOT move or delete originals.

Preserved:
- All original dated assessment files (audit trail)
- All historical versions (version control)
- File metadata and properties

Created:
- Normalized copies (for dashboard reading)
- Normalization manifest (audit documentation)

**Audit Trail:**

The normalization manifest documents:
- Which files were normalized
- When normalization occurred
- Source file → normalized file mapping
- File modification timestamps

Keep manifests for audit evidence of systematic quality control processes.

**Multiple Versions Handling:**

If multiple versions of same assessment exist (e.g., *_20260110.xlsx and
*_20260115.xlsx), normalization uses the MOST RECENT by file modification date.

This ensures dashboard always reads latest data without manual selection.

**Validation Before Normalization:**

Script validates each workbook has "Instructions & Legend" sheet before copying.
This prevents normalizing corrupt or incomplete workbooks.

Failed validations are reported but don't halt normalization of other files.

**File System Considerations:**

Normalized files created in SAME directory as source files because:
- Dashboard expects source files in same directory
- Keeps all IAM assessment files together organizationally
- Simplifies backup and version control

Don't split files across multiple directories - dashboard links will break.

**Re-Normalization Strategy:**

Safe to re-run normalization multiple times:
- Overwrites existing normalized files (desired behavior)
- Original dated files never touched (preservation)
- Each run creates new manifest with timestamp

Recommended re-normalization triggers:
- After any assessment workbook regeneration
- Before dashboard generation/regeneration
- After significant assessment data updates
- Monthly/quarterly during active development

**Quality Assurance:**

After normalization, validate:
```bash
# 1. Check normalized files exist
ls -lh ISMS-IMP-A.5.15-16-18.*.xlsx

# 2. Verify file sizes reasonable (not empty/corrupt)
# Typical sizes: 50KB-500KB per workbook

# 3. Test workbook opens correctly
# Open in Excel/LibreOffice and verify structure

# 4. Validate with sanity check script
python3 excel_sanity_check_a515.py ISMS-IMP-A.5.15-16-18.S1.xlsx
```

**Common Issues and Solutions:**

**Issue:** "No assessment workbooks found"
**Solution:** Ensure files match expected naming pattern: `ISMS-IMP-A.5.15-16-18.*_YYYYMMDD.xlsx`

**Issue:** "Workbook validation failed"
**Solution:** Source file corrupt or missing Instructions sheet. Regenerate assessment.

**Issue:** "Dashboard shows #REF! errors after normalization"
**Solution:** Dashboard and normalized files not in same directory. Move together.

**Issue:** "Normalized file exists but data is old"
**Solution:** Re-run normalization to update normalized copy from latest dated file.

**Issue:** "openpyxl not installed"
**Solution:** Run `sudo apt install python3-openpyxl` or `pip3 install openpyxl --break-system-packages`

================================================================================
"""

import os
import sys
import shutil
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("ℹ️  Install with: pip install openpyxl --break-system-packages")
    sys.exit(1)


# ============================================================================
# CONFIGURATION
# ============================================================================

# Expected document IDs and titles from assessment workbooks
EXPECTED_DOCS = {
    "ISMS-IMP-A.5.15-16-18.S1": {
        "title": "User Inventory & Lifecycle Compliance",
        "normalized": "ISMS-IMP-A.5.15-16-18.S1.xlsx"
    },
    "ISMS-IMP-A.5.15-16-18.S2": {
        "title": "Access Rights Matrix",
        "normalized": "ISMS-IMP-A.5.15-16-18.S2.xlsx"
    },
    "ISMS-IMP-A.5.15-16-18.S3": {
        "title": "Access Review Results",
        "normalized": "ISMS-IMP-A.5.15-16-18.S3.xlsx"
    },
    "ISMS-IMP-A.5.15-16-18.S4": {
        "title": "Role Compliance & SoD",
        "normalized": "ISMS-IMP-A.5.15-16-18.S4.xlsx"
    },
    "ISMS-IMP-A.5.15-16-18.S5": {
        "title": "Lifecycle Compliance Detailed",
        "normalized": "ISMS-IMP-A.5.15-16-18.S5.xlsx"
    },
}


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_workbook(filepath):
    """
    Validate workbook contains expected document ID in Instructions & Legend sheet.
    
    Args:
        filepath: Path to Excel workbook
    
    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        # Open workbook in read-only mode
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # Check for Instructions & Legend sheet
        sheet_name = None
        if "Instructions & Legend" in wb.sheetnames:
            sheet_name = "Instructions & Legend"
        elif "Instructions" in wb.sheetnames:
            sheet_name = "Instructions"
        else:
            # Try first sheet
            sheet_name = wb.sheetnames[0] if wb.sheetnames else None
        
        if not sheet_name:
            wb.close()
            return (None, None)
        
        ws = wb[sheet_name]
        
        # Look for Document ID in column A (typically rows 3-20)
        # Format: "Document ID:" label in column A, value in column B
        for row in range(3, 25):
            cell_label = ws.cell(row=row, column=1).value
            
            if cell_label and "Document ID" in str(cell_label):
                doc_id_value = ws.cell(row=row, column=2).value
                
                if doc_id_value:
                    doc_id = str(doc_id_value).strip()
                    
                    # Check if it's one of our expected documents
                    if doc_id in EXPECTED_DOCS:
                        title = EXPECTED_DOCS[doc_id]["title"]
                        wb.close()
                        return (doc_id, title)
        
        wb.close()
        return (None, None)
        
    except Exception as e:
        print(f"    ⚠️  Error reading file: {e}")
        return (None, None)


def get_file_info(filepath):
    """Get file metadata for manifest."""
    stat = filepath.stat()
    return {
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime),
        "created": datetime.fromtimestamp(stat.st_ctime),
    }


# ============================================================================
# SCANNING FUNCTIONS
# ============================================================================

def scan_directory(directory):
    """
    Scan directory for assessment workbooks.
    
    Args:
        directory: Path to scan
    
    Returns:
        dict: {doc_id: {'path': Path, 'info': metadata}}
    """
    found_assessments = {}
    directory = Path(directory)
    
    # Get all .xlsx files, excluding temporary files
    xlsx_files = [f for f in directory.glob("*.xlsx") if not f.name.startswith("~$")]
    
    if not xlsx_files:
        print(f"\n❌ No Excel files found in {directory}\n")
        return found_assessments
    
    print(f"\n🔍 Scanning {len(xlsx_files)} Excel file(s) in {directory}...\n")
    
    for filepath in sorted(xlsx_files):
        print(f"  Checking: {filepath.name}")
        doc_id, title = validate_workbook(filepath)
        
        if doc_id:
            print(f"    ✅ Valid: {doc_id} - {title}")
            
            # Check for duplicates
            if doc_id in found_assessments:
                print(f"\n    ⚠️  WARNING: DUPLICATE FOUND FOR {doc_id}")
                print(f"        Previous: {found_assessments[doc_id]['path'].name}")
                print(f"        Current:  {filepath.name}\n")
                
                response = input("        Use CURRENT file? (y/n): ").strip().lower()
                
                if response == 'y':
                    found_assessments[doc_id] = {
                        'path': filepath,
                        'info': get_file_info(filepath)
                    }
                    print(f"        → Replaced with current file\n")
                else:
                    print(f"        → Keeping previous file\n")
            else:
                found_assessments[doc_id] = {
                    'path': filepath,
                    'info': get_file_info(filepath)
                }
        else:
            print(f"    ⏭️  Skipped (not a valid assessment workbook)")
        
        print()  # Blank line for readability
    
    return found_assessments


# ============================================================================
# MANIFEST CREATION
# ============================================================================

def create_manifest(output_dir, mapping, source_dir):
    """
    Create audit manifest documenting the normalization process.
    
    Args:
        output_dir: Destination directory
        mapping: Dictionary of normalized files
        source_dir: Source directory
    
    Returns:
        Path: Path to created manifest file
    """
    manifest_path = Path(output_dir) / f"normalization_manifest_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    
    with open(manifest_path, 'w', encoding='utf-8') as f:
        # Header
        f.write("=" * 80 + "\n")
        f.write("ISMS IAM ASSESSMENT NORMALIZATION MANIFEST\n")
        f.write("ISO/IEC 27001:2022 - Controls A.5.15, A.5.16, A.5.18 (IAM Framework)\n")
        f.write("=" * 80 + "\n\n")
        
        # Normalization metadata
        f.write("NORMALIZATION METADATA\n")
        f.write("-" * 80 + "\n")
        f.write(f"Normalization Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Source Directory:        {source_dir.resolve()}\n")
        f.write(f"Output Directory:        {output_dir.resolve()}\n")
        f.write(f"Files Normalized:        {len(mapping)}/5 required\n")
        f.write(f"Normalization Status:    {'COMPLETE' if len(mapping) == 5 else 'INCOMPLETE'}\n")
        f.write("\n")
        
        # File mapping details
        f.write("FILE MAPPING\n")
        f.write("-" * 80 + "\n\n")
        
        for doc_id in sorted(EXPECTED_DOCS.keys()):
            if doc_id in mapping:
                info = mapping[doc_id]
                source_path = info['path']
                normalized_name = info['normalized']
                file_info = info['info']
                
                f.write(f"Document ID:         {doc_id}\n")
                f.write(f"Assessment Title:    {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"Source Filename:     {source_path.name}\n")
                f.write(f"Normalized Filename: {normalized_name}\n")
                f.write(f"Source File Size:    {file_info['size']:,} bytes\n")
                f.write(f"Source Modified:     {file_info['modified'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Source Created:      {file_info['created'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Status:              ✅ NORMALIZED\n")
            else:
                f.write(f"Document ID:         {doc_id}\n")
                f.write(f"Assessment Title:    {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"Status:              ❌ NOT FOUND\n")
            
            f.write("\n")
        
        # Summary
        f.write("-" * 80 + "\n")
        f.write("NORMALIZATION SUMMARY\n")
        f.write("-" * 80 + "\n\n")
        
        found_count = len(mapping)
        required_count = len(EXPECTED_DOCS)
        
        f.write(f"Assessment Coverage: {found_count}/{required_count} ({found_count/required_count*100:.0f}%)\n\n")
        
        if found_count == required_count:
            f.write("✅ All required assessment workbooks found and normalized.\n")
            f.write("   Dashboard can be generated with complete data linkage.\n")
        else:
            f.write(f"⚠️  WARNING: {required_count - found_count} assessment(s) missing.\n")
            f.write("   Dashboard will have incomplete data for missing assessments.\n")
            f.write("\n   Missing assessments:\n")
            for doc_id in EXPECTED_DOCS:
                if doc_id not in mapping:
                    f.write(f"   - {doc_id}: {EXPECTED_DOCS[doc_id]['title']}\n")
        
        f.write("\n")
        
        # Audit trail notes
        f.write("-" * 80 + "\n")
        f.write("AUDIT TRAIL NOTES\n")
        f.write("-" * 80 + "\n\n")
        f.write("Purpose:\n")
        f.write("  This normalization process prepares completed IAM assessment workbooks for\n")
        f.write("  integration and audit documentation.\n\n")
        f.write("Process:\n")
        f.write("  1. Source workbooks validated for correct Document ID\n")
        f.write("  2. Files copied (not moved) to preserve originals\n")
        f.write("  3. Normalized filenames applied (Document ID only, no dates/versions)\n")
        f.write("  4. Manifest created for audit traceability\n\n")
        f.write("Dashboard Integration:\n")
        f.write("  - Normalized files ready for dashboard linking\n")
        f.write("  - Place dashboard in same directory as normalized files\n")
        f.write("  - All files ready for compliance documentation\n\n")
        f.write("File Retention:\n")
        f.write("  - Original source files: Retained in source directory\n")
        f.write("  - Normalized copies: Located in output directory\n")
        f.write("  - This manifest: Retained with normalized files for audit\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("END OF MANIFEST\n")
        f.write("=" * 80 + "\n")
    
    return manifest_path


# ============================================================================
# MAIN NORMALIZATION FUNCTION
# ============================================================================

def normalize_files(source_dir=None, output_dir=None, auto_confirm=False):
    """
    Main normalization orchestration function.
    
    Args:
        source_dir: Source directory (prompts if None)
        output_dir: Output directory (prompts if None)
        auto_confirm: Skip confirmation prompts (for automation)
    
    Returns:
        bool: True if successful, False otherwise
    """
    print("=" * 80)
    print("ISMS IAM ASSESSMENT NORMALIZATION UTILITY")
    print("ISO/IEC 27001:2022 - Controls A.5.15, A.5.16, A.5.18 (IAM Framework)")
    print("=" * 80)
    print("\nThis script prepares assessment workbooks for dashboard integration by:")
    print("  1. Scanning for completed assessment files")
    print("  2. Validating document IDs in Instructions & Legend sheet")
    print("  3. Copying to normalized filenames (Document ID only, no dates)")
    print("  4. Creating audit manifest for traceability\n")
    
    # Get source directory
    if not source_dir:
        source_dir = input("Enter source directory path (or press Enter for current directory): ").strip()
        if not source_dir:
            source_dir = "."
    
    source_dir = Path(source_dir).resolve()
    
    if not source_dir.exists():
        print(f"\n❌ Error: Source directory does not exist: {source_dir}\n")
        return False
    
    if not source_dir.is_dir():
        print(f"\n❌ Error: Not a directory: {source_dir}\n")
        return False
    
    print(f"\n📁 Source directory: {source_dir}")
    
    # Get output directory
    if not output_dir:
        default_output = source_dir / "Dashboard_Sources"
        output_input = input(f"\nEnter output directory (or press Enter for '{default_output.name}'): ").strip()
        if not output_input:
            output_dir = default_output
        else:
            output_dir = Path(output_input)
    
    output_dir = Path(output_dir).resolve()
    
    # Create output directory if it doesn't exist
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        print(f"📁 Output directory: {output_dir}")
    except Exception as e:
        print(f"\n❌ Error creating output directory: {e}\n")
        return False
    
    # Scan for assessment files
    found = scan_directory(source_dir)
    
    if not found:
        print("❌ No valid assessment workbooks found in source directory\n")
        print("   Ensure files contain valid Document IDs in 'Instructions & Legend' sheet:")
        for doc_id, info in EXPECTED_DOCS.items():
            print(f"   - {doc_id}: {info['title']}")
        print()
        return False
    
    # Show normalization summary
    print("=" * 80)
    print("NORMALIZATION SUMMARY")
    print("=" * 80 + "\n")
    print(f"Found {len(found)} of {len(EXPECTED_DOCS)} required assessment workbooks:\n")
    
    mapping = {}
    for doc_id in sorted(EXPECTED_DOCS.keys()):
        if doc_id in found:
            normalized_name = EXPECTED_DOCS[doc_id]["normalized"]
            source_path = found[doc_id]['path']
            
            print(f"  ✅ {doc_id}")
            print(f"     Title:      {EXPECTED_DOCS[doc_id]['title']}")
            print(f"     Source:     {source_path.name}")
            print(f"     Normalized: {normalized_name}\n")
            
            mapping[doc_id] = {
                'path': source_path,
                'normalized': normalized_name,
                'info': found[doc_id]['info']
            }
        else:
            print(f"  ❌ {doc_id} - NOT FOUND")
            print(f"     Title: {EXPECTED_DOCS[doc_id]['title']}\n")
    
    if len(found) < len(EXPECTED_DOCS):
        print(f"⚠️  WARNING: Only {len(found)}/{len(EXPECTED_DOCS)} assessment workbooks found")
        print("   Dashboard will have incomplete data for missing assessments\n")
    
    print(f"Output directory: {output_dir}\n")
    
    # Confirm normalization
    if not auto_confirm:
        response = input("Proceed with normalization? (y/n): ").strip().lower()
        if response != 'y':
            print("\n❌ Normalization cancelled by user\n")
            return False
    
    # Perform normalization (copy files)
    print("\n" + "=" * 80)
    print("NORMALIZING FILES...")
    print("=" * 80 + "\n")
    
    for doc_id, info in mapping.items():
        source = info['path']
        dest = output_dir / info['normalized']
        
        print(f"Copying: {source.name}")
        print(f"      → {dest.name}")
        
        try:
            shutil.copy2(source, dest)
            print(f"      ✅ Success\n")
        except Exception as e:
            print(f"      ❌ Error: {e}\n")
            print(f"❌ Normalization failed at {doc_id}\n")
            return False
    
    # Create audit manifest
    print("📄 Creating audit manifest...")
    try:
        manifest = create_manifest(output_dir, mapping, source_dir)
        print(f"   ✅ Created: {manifest.name}\n")
    except Exception as e:
        print(f"   ❌ Error creating manifest: {e}\n")
        return False
    
    # Success summary
    print("=" * 80)
    print("✅ NORMALIZATION COMPLETE")
    print("=" * 80 + "\n")
    print(f"Normalized files:  {output_dir}")
    print(f"Audit manifest:    {manifest}\n")
    print("NEXT STEPS:\n")
    print("  1. Review audit manifest for file mapping details")
    print("  2. All files ready for compliance documentation and audit\n")
    print("  3. Place dashboard in same directory as normalized files:")
    print(f"     {output_dir}\n")
    print("=" * 80 + "\n")
    
    return True


# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Normalize ISMS IAM assessment workbooks",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode (prompts for directories)
  python3 normalize_assessment_files_a515-16-18.py
  
  # Specify source directory
  python3 normalize_assessment_files_a515-16-18.py --source ./assessments
  
  # Specify both directories
  python3 normalize_assessment_files_a515-16-18.py --source ./assessments --output ./dashboard
  
  # Automated mode (no prompts)
  python3 normalize_assessment_files_a515-16-18.py --source ./assessments --auto-confirm
        """
    )
    
    parser.add_argument(
        '--source', '-s',
        help='Source directory containing assessment workbooks (default: current directory)'
    )
    
    parser.add_argument(
        '--output', '-o',
        help='Output directory for normalized files (default: ./Dashboard_Sources)'
    )
    
    parser.add_argument(
        '--auto-confirm', '-y',
        action='store_true',
        help='Skip confirmation prompts (for automation)'
    )
    
    args = parser.parse_args()
    
    success = normalize_files(
        source_dir=args.source,
        output_dir=args.output,
        auto_confirm=args.auto_confirm
    )
    
    sys.exit(0 if success else 1)
