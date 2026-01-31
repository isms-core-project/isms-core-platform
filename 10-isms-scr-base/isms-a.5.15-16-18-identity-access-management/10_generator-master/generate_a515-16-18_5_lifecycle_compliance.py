#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.5.15-16-18.S5 - Lifecycle Compliance Detailed Assessment Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.16: Identity Management
Assessment Workbook 5 of 5: Detailed Identity Lifecycle Compliance Analysis

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific identity lifecycle processes, HR integration,
and assessment requirements.

Key customization areas:
1. Lifecycle event types and workflows (match your HR/identity integration)
2. Provisioning/deprovisioning SLA thresholds (aligned with your IAM policy)
3. Exception criteria and approval processes (based on your governance)
4. Metric calculation formulas (adapted to your compliance targets)
5. Evidence requirements (specific to your audit framework)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.15-16-18 IAM Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for detailed
analysis of identity lifecycle compliance across all joiner/mover/leaver events.

**Purpose:**
Enables granular assessment of identity lifecycle processes against
ISO 27001:2022 Control A.5.16 requirements, supporting evidence-based
validation of provisioning/deprovisioning timeliness and process effectiveness.

**Assessment Scope:**
- Joiner event detailed analysis (provisioning timeliness)
- Mover event detailed analysis (access modification timeliness)
- Leaver event detailed analysis (deprovisioning timeliness)
- Exception handling and approval tracking
- Process adherence metrics (% on-time vs. late)
- Root cause analysis for delays
- HR system integration effectiveness
- Automation vs. manual process tracking
- Lifecycle event volume and trends
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Usage guidance and lifecycle standards
2. Joiner_Events - Detailed provisioning events (60 joiners)
3. Mover_Events - Detailed role change/transfer events (40 movers)
4. Leaver_Events - Detailed deprovisioning events (50 leavers)
5. Exception_Tracking - Lifecycle exceptions and approvals (20 exceptions)
6. Process_Metrics - Detailed lifecycle compliance metrics
7. Root_Cause_Analysis - Delay reasons and improvement actions
8. Automation_Coverage - Automated vs. manual lifecycle tracking
9. Gap_Analysis - Process gaps and remediation tracking
10. Evidence_Register - Evidence collection for A.5.16 compliance
11. Approval_Sign_Off - 3-level approval workflow

**Key Features:**
- Data validation with lifecycle event type dropdown lists
- Conditional formatting for timeliness compliance (on-time/late/very late)
- Automated gap identification for delayed lifecycle events
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with HR system event exports

**Integration:**
This assessment feeds into the IAM Governance Dashboard, which
consolidates data from all five IAM assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - random (standard library - for sample data generation)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a515-16-18_5_lifecycle_compliance.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a515-16-18_5_lifecycle_compliance.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a515-16-18_5_lifecycle_compliance.py --date 20250124

Output:
    File: ISMS-IMP-A.5.15-16-18.S5_Lifecycle_Compliance_Detailed_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Export joiner/mover/leaver events from HR system
    2. Export corresponding identity system events (account creation/changes)
    3. Document each lifecycle event with timestamps
    4. Calculate timeliness (actual vs. SLA)
    5. Identify exceptions and document approvals
    6. Analyze root causes for delays
    7. Identify manual vs. automated processes
    8. Calculate process compliance metrics
    9. Identify improvement opportunities
    10. Collect and link audit evidence
    11. Obtain stakeholder approvals
    12. Feed results into IAM Governance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.16
Assessment Domain:    5 of 5 (Lifecycle Compliance Detailed Analysis)
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.15-16-18: Identity Management Policy (A.5.16)
    - ISMS-IMP-A.5.15-16-18-S2: Identity Lifecycle Process Guide
    - ISMS-IMP-A.5.15-16-18.S1: User Inventory Assessment (Workbook 1 - summary metrics)
    - ISMS-IMP-A.5.15-16-18.S2: Access Rights Matrix Assessment (Workbook 2)
    - ISMS-IMP-A.5.15-16-18.S3: Access Review Results Assessment (Workbook 3)
    - ISMS-IMP-A.5.15-16-18.S4: Role Compliance Assessment (Workbook 4)
    - IAM Governance Dashboard: Consolidated IAM compliance view

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.15-16-18-S2
    - Supports granular lifecycle event analysis and root cause investigation
    - Integrated with IAM Governance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Lifecycle Timeliness is CRITICAL:**
Identity lifecycle delays create security and business risks:
- Delayed joiner provisioning: New employee can't work (business impact)
- Delayed leaver deprovisioning: Ex-employee retains access (security risk)
- Delayed mover processing: Wrong access for role (compliance issue)

SLA compliance MUST be measured and monitored.

**SLA Standards:**
Typical lifecycle SLAs (adapt to your risk profile):
- Joiner provisioning: Access ready by start date (0 delay)
- Leaver deprovisioning: Access disabled within 24 hours of termination
- Mover processing: Access updated within 48 hours of role change
- Contractor expiration: Access auto-disabled on contract end date

Non-compliance = audit finding.

**Audit Considerations:**
Auditors expect to see:
- Documented lifecycle events with timestamps (HR event → identity event)
- Evidence of timely provisioning/deprovisioning (SLA compliance)
- Exception handling (documented approvals for delays)
- Root cause analysis (why delays occurred)
- Process improvements (remediation actions for systemic issues)

**Data Protection:**
Assessment workbooks contain sensitive employee lifecycle data:
- Employee names, hire/termination dates
- Organizational changes (transfers, role changes)
- Process performance metrics

Handle in accordance with GDPR/FADP data protection requirements.

**Maintenance:**
Review and update assessment:
- Monthly: Calculate lifecycle compliance metrics for latest events
- Quarterly: Analyze trends and identify systemic delays
- Semi-annually: Update SLA thresholds based on risk reassessment
- Annually: Complete lifecycle process effectiveness review
- Ad-hoc: When HR or identity system changes occur

**Quality Assurance:**
Lifecycle compliance depends on:
- Accurate HR data (authoritative source of truth)
- Reliable HR-to-identity integration (automated event triggers)
- Timestamp precision (prove compliance within SLA)
- Complete event logging (no missing lifecycle events)

Validate data accuracy through sampling and reconciliation.

**Automation is Key:**
Manual lifecycle processes = high delay risk:
- Email-based requests prone to delays
- Manual account creation error-prone
- No automated SLA tracking

Automated HR-to-identity integration = reliable compliance:
- New hire in HR → auto-triggers account creation
- Termination in HR → auto-disables account
- Role change in HR → auto-triggers access review
- Contractor end date → auto-expires access

Measure automation coverage and drive improvement.

**Root Cause Analysis:**
Common lifecycle delay causes:
- Incomplete HR data (missing start date, department, manager)
- Manual approval bottlenecks (approver on vacation, unresponsive)
- System integration failures (HR-to-identity sync broken)
- Missing automation (manual process steps)
- Inadequate staffing (IT overwhelmed with requests)

Document root causes and implement corrective actions.

**Exception Management:**
Not all delays are non-compliance:
- Approved exceptions (start date changed, delayed onboarding)
- Business justification (contractor extension approved)
- Technical dependencies (system maintenance, integration outage)

Exceptions MUST be:
- Documented with business justification
- Approved by appropriate authority (manager, security)
- Tracked separately (don't hide in metrics)
- Time-limited (exceptions reviewed periodically)

Unexplained delays = non-compliance.

**Continuous Improvement:**
Use lifecycle assessment to drive process improvements:
- High manual effort → automate
- Frequent delays → strengthen HR integration
- Systemic bottlenecks → redesign approval workflow
- Low compliance → increase monitoring and accountability

Track improvement trends quarter-over-quarter.

================================================================================
"""

import sys
from datetime import datetime, timedelta
from pathlib import Path
import random

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import PieChart, BarChart, Reference
except ImportError as e:
    print(f"❌ ERROR: Required library not found: {e}")
    print("📦 Install required libraries: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# SECTION 1: CONSTANTS AND CONFIGURATION
# ============================================================================

WORKBOOK_NAME = "Lifecycle Compliance Detailed Assessment"
DOCUMENT_ID = "ISMS-IMP-A.5.15-16-18.S5"
CONTROL_ID = "A.5.16"
CONTROL_NAME = "Identity Management"
GENERATED_DATE = datetime.now().strftime("%Y-%m-%d")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Data row counts
JOINER_EVENTS = 40
MOVER_EVENTS = 20
LEAVER_EVENTS = 30
CONTRACTOR_COUNT = 25
ORPHANED_REMEDIATION = 15
GAP_ROW_COUNT = 25
EVIDENCE_ROW_COUNT = 40

# SLA thresholds
PROVISIONING_SLA_DAYS = 0      # Provision by start date
DEPROVISIONING_SLA_HOURS = 24  # Deprovision within 24 hours


# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================

def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="002060", end_color="002060", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data_center": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "compliant": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "non_compliant": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "warning": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }
    
    return styles


def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = style_dict["font"]
    if "fill" in style_dict:
        cell.fill = style_dict["fill"]
    if "alignment" in style_dict:
        cell.alignment = style_dict["alignment"]
    if "border" in style_dict:
        cell.border = style_dict["border"]


# ============================================================================
# SECTION 3: SAMPLE DATA GENERATORS
# ============================================================================

def generate_sample_users(count, user_type="employee"):
    """Generate sample user data."""
    first_names = ["Alice", "Bob", "Carol", "David", "Emma", "Frank", "Grace", "Henry"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia"]
    departments = ["Finance", "Sales", "Marketing", "IT", "HR", "Operations", "Engineering"]
    
    users = []
    for i in range(count):
        first = random.choice(first_names)
        last = random.choice(last_names)
        username = f"{first.lower()}.{last.lower()}"
        if user_type == "contractor":
            username = f"EXT-{username}"
        
        users.append({
            "user_id": f"U{10000 + i}",
            "username": username,
            "full_name": f"{first} {last}",
            "department": random.choice(departments),
            "user_type": user_type.capitalize(),
        })
    
    return users


# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS
# ============================================================================

def create_instructions_sheet(wb, styles):
    """Create Sheet 1: Instructions & Legend."""
    ws = wb.create_sheet("Instructions & Legend", 0)
    
    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = f"{WORKBOOK_NAME} - Instructions"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 30
    
    # Document metadata
    ws["A3"].value = "Document ID:"
    ws["A3"].font = Font(bold=True)
    ws["B3"].value = DOCUMENT_ID
    
    ws["A4"].value = "Assessment:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].value = "Lifecycle Compliance Detailed"
    
    ws["A5"].value = "Version:"
    ws["A5"].font = Font(bold=True)
    ws["B5"].value = "1.0"
    
    ws["A6"].value = "Generated:"
    ws["A6"].font = Font(bold=True)
    ws["B6"].value = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    # Document info
    ws.merge_cells("A8:H8")
    cell = ws["A8"]
    cell.value = f"ISO/IEC 27001:2022 Control {CONTROL_ID}: {CONTROL_NAME}"
    apply_style(cell, styles["header"])
    
    ws.merge_cells("A9:H9")
    ws["A9"].value = f"Control Assessment Framework"
    ws["A9"].font = Font(name="Calibri", size=10, italic=True)
    
    # Purpose section
    row = 11
    ws[f"A{row}"] = "Purpose"
    apply_style(ws[f"A{row}"], styles["subheader"])
    ws.merge_cells(f"A{row}:H{row}")
    
    purpose_text = """
This workbook provides detailed identity lifecycle compliance tracking with event-level 
analysis. It monitors:
• Joiner events (new hire provisioning timeliness)
• Mover events (role change/transfer processing)
• Leaver events (termination deprovisioning compliance)
• Contractor lifecycle (time-bound access expiration)
• Orphaned account remediation progress
• HR system integration effectiveness
    """.strip()
    
    row += 1
    ws.merge_cells(f"A{row}:H{row+6}")
    ws[f"A{row}"].value = purpose_text
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 120
    
    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 35
    
    return ws


def create_joiner_events_sheet(wb, styles):
    """Create Sheet 2: Joiner_Events."""
    ws = wb.create_sheet("Joiner_Events")
    
    # Title
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = "Joiner Events - New Hire Provisioning Compliance"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Assessment Period: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Total Joiner Events: {JOINER_EVENTS}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    row = 5
    headers = [
        "Event ID", "User ID", "Username", "Full Name", "Department", "Job Title",
        "Hire Date", "Provision Date", "Days to Provision", "SLA", "Compliance", "Notes"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Generate events
    users = generate_sample_users(JOINER_EVENTS, "employee")
    job_titles = ["Analyst", "Manager", "Specialist", "Coordinator", "Engineer", "Representative"]
    
    for i, user in enumerate(users):
        row += 1
        hire_date = datetime.now() - timedelta(days=random.randint(30, 180))
        
        # Most compliant, some late
        if random.random() > 0.15:  # 85% compliant
            provision_date = hire_date - timedelta(days=random.randint(0, 2))
        else:  # 15% late
            provision_date = hire_date + timedelta(days=random.randint(1, 7))
        
        days_to_provision = (provision_date - hire_date).days
        compliant = days_to_provision <= PROVISIONING_SLA_DAYS
        
        ws.cell(row=row, column=1).value = f"JOIN-{1000 + i}"
        ws.cell(row=row, column=2).value = user["user_id"]
        ws.cell(row=row, column=3).value = user["username"]
        ws.cell(row=row, column=4).value = user["full_name"]
        ws.cell(row=row, column=5).value = user["department"]
        ws.cell(row=row, column=6).value = random.choice(job_titles)
        ws.cell(row=row, column=7).value = hire_date.strftime("%Y-%m-%d")
        ws.cell(row=row, column=8).value = provision_date.strftime("%Y-%m-%d")
        ws.cell(row=row, column=9).value = days_to_provision
        ws.cell(row=row, column=10).value = PROVISIONING_SLA_DAYS
        ws.cell(row=row, column=11).value = "Compliant" if compliant else "Non-Compliant"
        
        if not compliant:
            ws.cell(row=row, column=12).value = f"Provisioned {abs(days_to_provision)} day(s) late"
        
        # Formatting
        comp_cell = ws.cell(row=row, column=11)
        if compliant:
            apply_style(comp_cell, styles["compliant"])
        else:
            apply_style(comp_cell, styles["non_compliant"])
        
        for col in range(1, 13):
            if col != 11:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 12, "B": 10, "C": 18, "D": 20, "E": 15, "F": 18, "G": 12, "H": 15, "I": 15, "J": 8, "K": 15, "L": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_mover_events_sheet(wb, styles):
    """Create Sheet 3: Mover_Events."""
    ws = wb.create_sheet("Mover_Events")
    
    # Title
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = "Mover Events - Role Change & Transfer Processing"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Assessment Period: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Total Mover Events: {MOVER_EVENTS}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    row = 5
    headers = [
        "Event ID", "User ID", "Username", "Full Name", "Change Type",
        "From Dept", "To Dept", "Effective Date", "Access Updated", "Days to Update", "Compliance", "Notes"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Generate events
    users = generate_sample_users(MOVER_EVENTS, "employee")
    departments = ["Finance", "Sales", "Marketing", "IT", "HR", "Operations", "Engineering"]
    change_types = ["Department Transfer", "Promotion", "Role Change", "Team Reassignment"]
    
    for i, user in enumerate(users):
        row += 1
        effective_date = datetime.now() - timedelta(days=random.randint(7, 90))
        
        # Most compliant, some delayed
        if random.random() > 0.2:  # 80% compliant
            update_date = effective_date + timedelta(days=random.randint(0, 3))
        else:
            update_date = effective_date + timedelta(days=random.randint(4, 14))
        
        days_to_update = (update_date - effective_date).days
        compliant = days_to_update <= 3  # 3-day SLA for movers
        
        from_dept = user["department"]
        to_dept = random.choice([d for d in departments if d != from_dept])
        
        ws.cell(row=row, column=1).value = f"MOVE-{1000 + i}"
        ws.cell(row=row, column=2).value = user["user_id"]
        ws.cell(row=row, column=3).value = user["username"]
        ws.cell(row=row, column=4).value = user["full_name"]
        ws.cell(row=row, column=5).value = random.choice(change_types)
        ws.cell(row=row, column=6).value = from_dept
        ws.cell(row=row, column=7).value = to_dept
        ws.cell(row=row, column=8).value = effective_date.strftime("%Y-%m-%d")
        ws.cell(row=row, column=9).value = update_date.strftime("%Y-%m-%d")
        ws.cell(row=row, column=10).value = days_to_update
        ws.cell(row=row, column=11).value = "Compliant" if compliant else "Delayed"
        
        if not compliant:
            ws.cell(row=row, column=12).value = f"Access update delayed {days_to_update} day(s)"
        
        # Formatting
        comp_cell = ws.cell(row=row, column=11)
        if compliant:
            apply_style(comp_cell, styles["compliant"])
        else:
            apply_style(comp_cell, styles["warning"])
        
        for col in range(1, 13):
            if col != 11:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 12, "B": 10, "C": 18, "D": 20, "E": 20, "F": 15, "G": 15, "H": 15, "I": 15, "J": 15, "K": 15, "L": 35}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_leaver_events_sheet(wb, styles):
    """Create Sheet 4: Leaver_Events."""
    ws = wb.create_sheet("Leaver_Events")
    
    # Title
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = "Leaver Events - Termination Deprovisioning Compliance"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Assessment Period: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Total Leaver Events: {LEAVER_EVENTS}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    row = 5
    headers = [
        "Event ID", "User ID", "Username", "Full Name", "Department",
        "Termination Date", "Account Disabled", "Hours to Disable", "SLA", "Compliance", "Security Risk", "Notes"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Generate events
    users = generate_sample_users(LEAVER_EVENTS, "employee")
    
    for i, user in enumerate(users):
        row += 1
        term_date = datetime.now() - timedelta(days=random.randint(1, 60))
        
        # Most compliant, some late (critical security issue)
        if random.random() > 0.1:  # 90% compliant
            disable_date = term_date + timedelta(hours=random.randint(0, 12))
        else:  # 10% late
            disable_date = term_date + timedelta(days=random.randint(2, 10))
        
        hours_to_disable = (disable_date - term_date).total_seconds() / 3600
        compliant = hours_to_disable <= DEPROVISIONING_SLA_HOURS
        
        ws.cell(row=row, column=1).value = f"LEAVE-{1000 + i}"
        ws.cell(row=row, column=2).value = user["user_id"]
        ws.cell(row=row, column=3).value = user["username"]
        ws.cell(row=row, column=4).value = user["full_name"]
        ws.cell(row=row, column=5).value = user["department"]
        ws.cell(row=row, column=6).value = term_date.strftime("%Y-%m-%d")
        ws.cell(row=row, column=7).value = disable_date.strftime("%Y-%m-%d %H:%M")
        ws.cell(row=row, column=8).value = round(hours_to_disable, 1)
        ws.cell(row=row, column=9).value = DEPROVISIONING_SLA_HOURS
        ws.cell(row=row, column=10).value = "Compliant" if compliant else "Non-Compliant"
        ws.cell(row=row, column=11).value = "No" if compliant else "YES"
        
        if not compliant:
            ws.cell(row=row, column=12).value = f"Terminated user retained access for {int(hours_to_disable/24)} day(s) - SECURITY RISK"
        
        # Formatting
        comp_cell = ws.cell(row=row, column=10)
        risk_cell = ws.cell(row=row, column=11)
        
        if compliant:
            apply_style(comp_cell, styles["compliant"])
            apply_style(risk_cell, styles["compliant"])
        else:
            apply_style(comp_cell, styles["non_compliant"])
            apply_style(risk_cell, styles["non_compliant"])
        
        for col in range(1, 13):
            if col not in [10, 11]:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 12, "B": 10, "C": 18, "D": 20, "E": 15, "F": 15, "G": 18, "H": 15, "I": 8, "J": 15, "K": 15, "L": 50}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_contractor_lifecycle_sheet(wb, styles):
    """Create Sheet 5: Contractor_Lifecycle."""
    ws = wb.create_sheet("Contractor_Lifecycle")
    
    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "Contractor Lifecycle - Time-Bound Access Management"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Active Contractors: {CONTRACTOR_COUNT}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    row = 5
    headers = [
        "Contractor ID", "Username", "Full Name", "Sponsor", "Department",
        "Contract Start", "Contract End", "Days Remaining", "Expiration Status", "Extension Approved", "Compliance"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Generate contractors
    contractors = generate_sample_users(CONTRACTOR_COUNT, "contractor")
    
    for contractor in contractors:
        row += 1
        start_date = datetime.now() - timedelta(days=random.randint(90, 365))
        end_date = datetime.now() + timedelta(days=random.randint(-30, 180))
        days_remaining = (end_date - datetime.now()).days
        
        ws.cell(row=row, column=1).value = contractor["user_id"]
        ws.cell(row=row, column=2).value = contractor["username"]
        ws.cell(row=row, column=3).value = contractor["full_name"]
        ws.cell(row=row, column=4).value = f"{contractor['department']} Manager"
        ws.cell(row=row, column=5).value = contractor["department"]
        ws.cell(row=row, column=6).value = start_date.strftime("%Y-%m-%d")
        ws.cell(row=row, column=7).value = end_date.strftime("%Y-%m-%d")
        ws.cell(row=row, column=8).value = days_remaining
        
        # Status
        if days_remaining < 0:
            status = "EXPIRED"
            extension = "N/A"
            compliance = "Non-Compliant"
            style = styles["non_compliant"]
        elif days_remaining <= 30:
            status = "Expiring Soon"
            extension = random.choice(["Approved", "Pending", "Not Requested"])
            compliance = "Warning"
            style = styles["warning"]
        else:
            status = "Active"
            extension = "N/A"
            compliance = "Compliant"
            style = styles["compliant"]
        
        ws.cell(row=row, column=9).value = status
        ws.cell(row=row, column=10).value = extension
        ws.cell(row=row, column=11).value = compliance
        
        apply_style(ws.cell(row=row, column=11), style)
        
        for col in range(1, 11):
            apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 12, "B": 22, "C": 20, "D": 20, "E": 15, "F": 15, "G": 15, "H": 15, "I": 18, "J": 18, "K": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_orphaned_remediation_sheet(wb, styles):
    """Create Sheet 6: Orphaned_Remediation."""
    ws = wb.create_sheet("Orphaned_Remediation")
    
    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "Orphaned Account Remediation Tracking"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True, bold=True, color="FF0000")
    ws["A3"] = f"Orphaned Accounts: {ORPHANED_REMEDIATION}"
    ws["A3"].font = Font(italic=True, bold=True, color="FF0000")
    
    # Column headers
    row = 5
    headers = [
        "Account ID", "Username", "Account Type", "Last Login", "Detection Date",
        "Orphan Reason", "Assigned To", "Remediation Action", "Completion Date", "Days to Remediate", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Generate orphaned accounts
    users = generate_sample_users(ORPHANED_REMEDIATION, "employee")
    reasons = [
        "No login for 120+ days",
        "User terminated but account not disabled",
        "No valid business owner",
        "Contractor contract ended, account still active",
    ]
    actions = ["Disable account", "Delete account", "Reassign owner", "Document exception"]
    
    for i, user in enumerate(users):
        row += 1
        detection_date = datetime.now() - timedelta(days=random.randint(1, 30))
        
        # Some remediated, some pending
        if random.random() > 0.4:
            completion_date = detection_date + timedelta(days=random.randint(1, 10))
            days_to_remediate = (completion_date - detection_date).days
            status = "Remediated"
            style = styles["compliant"]
        else:
            completion_date = None
            days_to_remediate = None
            status = "Pending"
            style = styles["non_compliant"]
        
        ws.cell(row=row, column=1).value = user["user_id"]
        ws.cell(row=row, column=2).value = user["username"]
        ws.cell(row=row, column=3).value = random.choice(["Employee", "Contractor", "Service Account"])
        ws.cell(row=row, column=4).value = (datetime.now() - timedelta(days=random.randint(90, 365))).strftime("%Y-%m-%d")
        ws.cell(row=row, column=5).value = detection_date.strftime("%Y-%m-%d")
        ws.cell(row=row, column=6).value = random.choice(reasons)
        ws.cell(row=row, column=7).value = random.choice(["IT Operations", "Security Team", "Department Manager"])
        ws.cell(row=row, column=8).value = random.choice(actions)
        ws.cell(row=row, column=9).value = completion_date.strftime("%Y-%m-%d") if completion_date else ""
        ws.cell(row=row, column=10).value = days_to_remediate if days_to_remediate else ""
        ws.cell(row=row, column=11).value = status
        
        apply_style(ws.cell(row=row, column=11), style)
        
        for col in range(1, 11):
            apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 10, "B": 18, "C": 15, "D": 15, "E": 15, "F": 45, "G": 20, "H": 22, "I": 15, "J": 18, "K": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_metrics_sheet(wb, styles):
    """Create Sheet 7: Timeliness_Metrics."""
    ws = wb.create_sheet("Timeliness_Metrics")
    
    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Lifecycle Timeliness Metrics - SLA Compliance"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Assessment Period: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Metrics
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "Lifecycle Compliance Metrics"
    apply_style(cell, styles["header"])
    
    row += 1
    headers = ["Metric", "Target", "Actual", "Status", "Gap", "Comments"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    metrics = [
        ("Joiner On-Time Provisioning", "≥95%", "85%", "Warning", "-10%", f"{JOINER_EVENTS * 0.15:.0f} late provisioning events"),
        ("Leaver On-Time Deprovisioning", "≥98%", "90%", "Warning", "-8%", f"{LEAVER_EVENTS * 0.10:.0f} late deprovisioning events"),
        ("Mover Access Update", "≥90%", "80%", "Warning", "-10%", f"{MOVER_EVENTS * 0.20:.0f} delayed updates"),
        ("Contractor Expiration Compliance", "100%", "88%", "Warning", "-12%", "3 expired contractor accounts"),
        ("Orphaned Account Remediation", "≥90%", "60%", "Non-Compliant", "-30%", f"{ORPHANED_REMEDIATION * 0.40:.0f} pending remediation"),
    ]
    
    for metric_name, target, actual, status, gap, comments in metrics:
        row += 1
        ws.cell(row=row, column=1).value = metric_name
        ws.cell(row=row, column=2).value = target
        ws.cell(row=row, column=3).value = actual
        ws.cell(row=row, column=4).value = status
        ws.cell(row=row, column=5).value = gap
        ws.cell(row=row, column=6).value = comments
        
        status_cell = ws.cell(row=row, column=4)
        if status == "Compliant":
            apply_style(status_cell, styles["compliant"])
        elif status == "Warning":
            apply_style(status_cell, styles["warning"])
        else:
            apply_style(status_cell, styles["non_compliant"])
        
        for col in range(1, 7):
            if col != 4:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 35, "B": 15, "C": 15, "D": 18, "E": 10, "F": 50}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    return ws


def create_hr_integration_sheet(wb, styles):
    """Create Sheet 8: HR_Integration."""
    ws = wb.create_sheet("HR_Integration")
    
    # Title
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "HR System Integration Status"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Integration status
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws[f"A{row}"]
    cell.value = "Integration Components"
    apply_style(cell, styles["header"])
    
    row += 1
    headers = ["Component", "Status", "Integration Method", "Frequency", "Last Sync", "Issues", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    components = [
        ("Employee Data Feed", "Active", "API", "Real-time", (datetime.now() - timedelta(hours=1)).strftime("%Y-%m-%d %H:%M"), "None", "Automated employee sync"),
        ("Termination Feed", "Partial", "Manual", "Daily", (datetime.now() - timedelta(hours=24)).strftime("%Y-%m-%d %H:%M"), "Manual delays", "Automation pending"),
        ("Department Transfer Feed", "Not Implemented", "N/A", "N/A", "N/A", "No automated feed", "Requires manual process"),
        ("Contractor Data Feed", "Active", "CSV Import", "Weekly", (datetime.now() - timedelta(days=3)).strftime("%Y-%m-%d %H:%M"), "None", "Weekly contractor sync"),
        ("Authoritative Source Validation", "Active", "API", "Hourly", (datetime.now() - timedelta(minutes=30)).strftime("%Y-%m-%d %H:%M"), "None", "HR is authoritative source"),
    ]
    
    for component_data in components:
        row += 1
        for col, value in enumerate(component_data, start=1):
            ws.cell(row=row, column=col).value = value
        
        # Status formatting
        status_cell = ws.cell(row=row, column=2)
        if component_data[1] == "Active":
            apply_style(status_cell, styles["compliant"])
        elif component_data[1] == "Partial":
            apply_style(status_cell, styles["warning"])
        else:
            apply_style(status_cell, styles["non_compliant"])
        
        for col in range(1, 8):
            if col != 2:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 35, "B": 18, "C": 20, "D": 15, "E": 18, "F": 20, "G": 35}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    return ws


def create_gap_analysis_sheet(wb, styles):
    """Create Sheet 9: Gap_Analysis."""
    ws = wb.create_sheet("Gap_Analysis")
    
    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "Gap Analysis - Lifecycle Process Non-Compliance"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Column headers
    row = 4
    headers = [
        "Gap ID", "Category", "Description", "Risk Level", "Affected Items",
        "Root Cause", "Remediation Plan", "Owner", "Due Date", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Sample gaps
    gaps = [
        ("GAP-001", "Provisioning", "6 late provisioning events", "Medium", "6 events",
         "Manual provisioning process", "Automate provisioning via HR API", "IT Operations",
         (datetime.now() + timedelta(days=60)).strftime("%Y-%m-%d"), "Planned"),
        ("GAP-002", "Deprovisioning", "3 late deprovisioning events", "High", "3 events",
         "No automated termination feed", "Implement real-time termination feed", "Security Team",
         (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"), "In Progress"),
        ("GAP-003", "Orphaned Accounts", "6 pending orphaned account remediation", "High", "6 accounts",
         "No monthly automated scan", "Schedule monthly orphaned account scan", "Security Team",
         (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d"), "Open"),
        ("GAP-004", "Contractor Management", "3 expired contractor accounts", "Medium", "3 contractors",
         "No automated contract expiry alerts", "Implement contract expiry monitoring", "IAM Manager",
         (datetime.now() + timedelta(days=21)).strftime("%Y-%m-%d"), "Open"),
    ]
    
    for gap_data in gaps:
        row += 1
        for col, value in enumerate(gap_data, start=1):
            ws.cell(row=row, column=col).value = value
        
        # Color code by risk
        risk_cell = ws.cell(row=row, column=4)
        if gap_data[3] == "High":
            apply_style(risk_cell, styles["non_compliant"])
        elif gap_data[3] == "Medium":
            apply_style(risk_cell, styles["warning"])
        else:
            apply_style(risk_cell, styles["compliant"])
        
        # Color code status
        status_cell = ws.cell(row=row, column=10)
        if gap_data[9] == "Closed":
            apply_style(status_cell, styles["compliant"])
        elif gap_data[9] in ["In Progress", "Planned"]:
            apply_style(status_cell, styles["warning"])
        else:
            apply_style(status_cell, styles["non_compliant"])
        
        for col in range(1, 11):
            if col not in [4, 10]:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 10, "B": 18, "C": 35, "D": 12, "E": 15, "F": 35, "G": 45, "H": 18, "I": 12, "J": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A5"
    
    return ws


def create_evidence_register_sheet(wb, styles):
    """Create Sheet 10: Evidence_Register."""
    ws = wb.create_sheet("Evidence_Register")
    
    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "Evidence Register - A.5.16 Identity Management"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = f"Evidence Collection Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Column headers
    row = 4
    headers = [
        "Evidence ID", "Requirement", "Evidence Type", "Evidence Location",
        "Collection Date", "Completeness", "Reviewed By", "Notes"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Sample evidence
    evidence_items = [
        ("EV-001", "Joiner Process", "Event Log", "Joiner_Events sheet", GENERATED_DATE, "Complete", "IAM Manager", f"{JOINER_EVENTS} joiner events tracked"),
        ("EV-002", "Mover Process", "Event Log", "Mover_Events sheet", GENERATED_DATE, "Complete", "IAM Manager", f"{MOVER_EVENTS} mover events tracked"),
        ("EV-003", "Leaver Process", "Event Log", "Leaver_Events sheet", GENERATED_DATE, "Complete", "IAM Manager", f"{LEAVER_EVENTS} leaver events tracked"),
        ("EV-004", "Contractor Lifecycle", "Tracking Sheet", "Contractor_Lifecycle sheet", GENERATED_DATE, "Complete", "IAM Manager", f"{CONTRACTOR_COUNT} contractors tracked"),
        ("EV-005", "Orphaned Remediation", "Remediation Log", "Orphaned_Remediation sheet", GENERATED_DATE, "Complete", "Security Team", f"{ORPHANED_REMEDIATION} orphaned accounts"),
        ("EV-006", "HR Integration", "Integration Status", "HR_Integration sheet", GENERATED_DATE, "Partial", "IT Manager", "Termination feed not automated"),
        ("EV-007", "Lifecycle Policy", "Policy Document", "ISMS-POL-A.5.15-16-18", GENERATED_DATE, "Complete", "CISO", "Policy approved"),
        ("EV-008", "Process Procedures", "Implementation Guide", "ISMS-IMP-A.5.15-16-18-S2", GENERATED_DATE, "Complete", "IAM Manager", "Procedures documented"),
    ]
    
    for evidence in evidence_items:
        row += 1
        for col, value in enumerate(evidence, start=1):
            ws.cell(row=row, column=col).value = value
        
        # Color code completeness
        completeness_cell = ws.cell(row=row, column=6)
        if evidence[5] == "Complete":
            apply_style(completeness_cell, styles["compliant"])
        elif evidence[5] == "Partial":
            apply_style(completeness_cell, styles["warning"])
        else:
            apply_style(completeness_cell, styles["non_compliant"])
        
        for col in range(1, 9):
            if col != 6:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Column widths
    widths = {"A": 12, "B": 25, "C": 20, "D": 35, "E": 18, "F": 15, "G": 20, "H": 45}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A5"
    
    return ws


# ============================================================================
# SECTION 5: MAIN EXECUTION
# ============================================================================

def main():
    """Main function to generate the workbook."""
    print("="*80)
    print(f"Generating: {WORKBOOK_NAME}")
    print(f"Control: ISO/IEC 27001:2022 {CONTROL_ID} - {CONTROL_NAME}")
    print("="*80)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Setup styles
    styles = setup_styles()
    
    # Create sheets
    print("\n📝 Creating sheets...")
    
    print("  1/10 Creating Instructions & Legend...")
    create_instructions_sheet(wb, styles)
    
    print("  2/10 Creating Joiner Events...")
    create_joiner_events_sheet(wb, styles)
    
    print("  3/10 Creating Mover Events...")
    create_mover_events_sheet(wb, styles)
    
    print("  4/10 Creating Leaver Events...")
    create_leaver_events_sheet(wb, styles)
    
    print("  5/10 Creating Contractor Lifecycle...")
    create_contractor_lifecycle_sheet(wb, styles)
    
    print("  6/10 Creating Orphaned Remediation...")
    create_orphaned_remediation_sheet(wb, styles)
    
    print("  7/10 Creating Timeliness Metrics...")
    create_metrics_sheet(wb, styles)
    
    print("  8/10 Creating HR Integration...")
    create_hr_integration_sheet(wb, styles)
    
    print("  9/10 Creating Gap Analysis...")
    create_gap_analysis_sheet(wb, styles)
    
    print("  10/10 Creating Evidence Register...")
    create_evidence_register_sheet(wb, styles)
    
    # Save workbook
    filename = f"Lifecycle_Compliance_Detailed_{GENERATED_TIMESTAMP}.xlsx"
    output_path = Path.cwd() / filename
    
    print(f"\n💾 Saving workbook: {filename}")
    wb.save(output_path)
    
    print("\n" + "="*80)
    print("✅ SUCCESS!")
    print(f"📄 File created: {output_path}")
    print(f"📊 Sheets: 10")
    print(f"📋 Lifecycle events: {JOINER_EVENTS + MOVER_EVENTS + LEAVER_EVENTS}")
    print("="*80)
    print("\nWorkbook Contents:")
    print(f"  • Joiner Events ({JOINER_EVENTS} provisioning events)")
    print(f"  • Mover Events ({MOVER_EVENTS} role change events)")
    print(f"  • Leaver Events ({LEAVER_EVENTS} termination events)")
    print(f"  • Contractor Lifecycle ({CONTRACTOR_COUNT} contractors)")
    print(f"  • Orphaned Account Remediation ({ORPHANED_REMEDIATION} accounts)")
    print("  • Timeliness Metrics & SLA Compliance")
    print("  • HR System Integration Status")
    print("  • Gap Analysis & Remediation")
    print("  • Evidence Register for A.5.16")
    print("="*80)
    
    return output_path


if __name__ == "__main__":
    try:
        output_file = main()
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
