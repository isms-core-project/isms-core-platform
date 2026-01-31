#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.23.5 - Compliance Summary Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.23: Web Filtering
Assessment Domain 5: Executive Compliance Dashboard & Consolidation

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific reporting requirements, stakeholder audiences,
and compliance scoring methodology.

Key customization areas:
1. Compliance scoring weights (match your risk-based prioritization)
2. Dashboard KPIs and metrics (adapt to stakeholder requirements)
3. Workbook file naming conventions (match your naming standards)
4. Data consolidation logic (specific to your assessment structures)
5. Reporting frequency and distribution (match your governance cadence)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

CRITICAL: This script consolidates data from the four A.8.23 assessment
workbooks. You MUST ensure that source workbook structures match the
expected schemas defined in the WORKBOOK_SCHEMAS configuration section.

Reference Pattern: Based on ISMS-A.8.23 Web Filtering Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive executive compliance dashboard that
consolidates assessment data from all four A.8.23 Web Filtering assessment
domains into a unified oversight workbook for governance and audit reporting.

**Purpose:**
Provides executive leadership, compliance teams, and auditors with a
consolidated view of web filtering control effectiveness across all assessment
domains, enabling risk-based decision making and audit readiness validation.

**Consolidation Scope:**
This dashboard consolidates data from four A.8.23 assessment workbooks:

1. **ISMS-IMP-A.8.23.1** - Filtering Infrastructure Assessment
   - Technology capabilities and infrastructure adequacy
   - Solution architecture and deployment validation
   - Integration effectiveness and performance metrics

2. **ISMS-IMP-A.8.23.2** - Network Coverage Assessment
   - Network segment filtering coverage verification
   - Device-level deployment validation
   - Coverage gap identification and exemption tracking

3. **ISMS-IMP-A.8.23.3** - Policy Configuration Assessment
   - Threat protection policy effectiveness
   - Category filtering and acceptable use alignment
   - Policy exception management and justification

4. **ISMS-IMP-A.8.23.4** - Monitoring & Response Assessment
   - Logging, monitoring, and alerting capabilities
   - Incident detection and response effectiveness
   - SIEM integration and reporting adequacy

**Generated Dashboard Structure:**
1. Executive Dashboard - High-level compliance summary and KPIs
2. Domain Overview - Detailed status by assessment domain
3. Gap Summary - Consolidated gaps requiring remediation
4. Compliance Trends - Historical compliance tracking (if available)
5. Evidence Summary - Audit evidence completeness verification
6. Action Items - Prioritized remediation and improvement actions
7. Risk Assessment - Residual risk evaluation and mitigation status
8. Audit Readiness - Stage-1 and Stage-2 audit preparation checklist
9. Data Sources - Source workbook inventory and validation status
10. Instructions & Methodology - Dashboard usage and interpretation guide

**Key Features:**
- **Automated Data Consolidation**: External workbook formula linking
- **Executive KPIs**: Overall compliance percentage, domain-level scores
- **Visual Dashboards**: Conditional formatting for at-a-glance status
- **Gap Prioritization**: Risk-based ranking of remediation requirements
- **Audit Readiness**: Evidence completeness and traceability verification
- **Trend Analysis**: Historical compliance tracking (if data available)
- **Action Tracking**: Remediation action register with ownership
- **Multi-Stakeholder View**: Tailored views for executives, auditors, operators

**Data Flow Architecture:**
```
Assessment Workbooks → Normalization → Dashboard Consolidation → Executive Reporting
     (Domains 1-4)         Script           (This Script)           & Audit Evidence
```

**Integration Requirements:**
1. Source workbooks MUST be normalized using `normalize_assessment_files_a823.py`
2. All four source workbooks MUST be in the same directory as dashboard
3. Dashboard file names MUST match the expected naming conventions
4. Excel "Update Links" must be enabled after opening dashboard
5. Source workbooks should not be modified after dashboard generation

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation
    - Completed and normalized A.8.23.1, A.8.23.2, A.8.23.3, A.8.23.4 workbooks

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

Prerequisites:
    1. Complete all four assessment workbooks (A.8.23.1 through A.8.23.4)
    2. Run normalization script: python3 normalize_assessment_files_a823.py
    3. Verify normalized files exist with correct naming

Expected Source Files:
    - ISMS_A_8_23_1_Filtering_Infrastructure_Assessment_YYYYMMDD.xlsx
    - ISMS_A_8_23_2_Network_Coverage_Assessment_YYYYMMDD.xlsx
    - ISMS_A_8_23_3_Policy_Configuration_Assessment_YYYYMMDD.xlsx
    - ISMS_A_8_23_4_Monitoring_Response_Assessment_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Workflow:
    1. Complete all four source assessment workbooks
    2. Normalize source files:
       python3 normalize_assessment_files_a823.py
    
    3. Generate dashboard:
       python3 generate_a823_5_compliance_dashboard.py
    
    4. Place dashboard in same directory as normalized source files
    5. Open dashboard in Excel
    6. Enable "Update Links" when prompted (CRITICAL!)
    7. Review consolidated compliance status

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a823_5_compliance_dashboard.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a823_5_compliance_dashboard.py --date 20250115
    
    # Generate with custom source file dates (for linking)
    python3 generate_a823_5_compliance_dashboard.py --source-date 20250110

Output:
    File: ISMS_A_8_23_5_Compliance_Dashboard_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Open dashboard in Excel (enable macros if prompted)
    2. **CRITICAL**: Enable "Update Links" when prompted
    3. Verify all data consolidation is working (no #REF! errors)
    4. Review executive dashboard for overall compliance status
    5. Investigate gaps and prioritize remediation actions
    6. Validate evidence completeness for audit readiness
    7. Review and approve action items with stakeholders
    8. Schedule regular dashboard updates (monthly/quarterly)
    9. Distribute to governance, compliance, and audit stakeholders
    10. Maintain version control and historical snapshots

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.23
Assessment Domain:    5 (Consolidation & Executive Oversight)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.23: Web Filtering Policy (Governance)
    - ISMS-IMP-A.8.23.1: Filtering Infrastructure Assessment (Domain 1)
    - ISMS-IMP-A.8.23.2: Network Coverage Assessment (Domain 2)
    - ISMS-IMP-A.8.23.3: Policy Configuration Assessment (Domain 3)
    - ISMS-IMP-A.8.23.4: Monitoring & Response Assessment (Domain 4)
    - ISMS-IMP-A.8.23.5: Compliance Dashboard Specification

Dependency Scripts:
    - normalize_assessment_files_a823.py (REQUIRED - run before this script)
    - excel_sanity_check_a823.py (optional - validation tool)
    - sanity_check_a823_dashboard.py (optional - dashboard validation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Consolidates all four A.8.23 assessment domains
    - Implements executive dashboard with KPIs and trend analysis
    - Provides audit readiness assessment and evidence tracking
    - Integrates with normalized assessment workbooks

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**External Workbook Formula Linking:**
This dashboard uses Excel external workbook references to consolidate data
from source assessment workbooks. This approach provides:
\u2705 Automatic updates when source data changes
\u2705 No manual data entry or copy-paste errors
\u2705 Real-time dashboard refresh capability
\u2705 Version control through linked source files

\u26A0\uFE0F CRITICAL Requirements:
- Source workbooks MUST exist in same directory as dashboard
- Source workbook file names MUST match expected conventions
- Users MUST enable "Update Links" when opening dashboard
- Moving source files breaks links (update paths if necessary)

**Audit Considerations:**
The compliance dashboard is a primary audit artifact. Ensure:
- All source assessments are complete and approved before generation
- Dashboard data accurately reflects source assessment status
- Evidence completeness is verified (not assumed)
- Gap remediation actions have owners and target dates
- Historical compliance trends demonstrate improvement
- Audit readiness checklist is complete before Stage-1 audit

Auditors will request:
- Source assessment workbooks with evidence linkage
- Gap analysis and remediation tracking
- Evidence of regular dashboard review (meeting minutes)
- Action item ownership and closure evidence
- Compliance trend analysis showing continuous improvement

**Data Protection:**
Dashboard workbooks contain comprehensive security control status information:
- Overall security posture and control effectiveness
- Identified gaps and vulnerabilities
- Remediation priorities and action plans
- Compliance status and audit findings

Handle in accordance with your organization's data classification policies
(typically CONFIDENTIAL or RESTRICTED).

**Maintenance:**
Update dashboard on regular cadence:
- Monthly: For active remediation projects
- Quarterly: For stable operations with periodic assessment
- After significant changes: Infrastructure, policy, or architecture changes
- Pre-audit: 2-4 weeks before scheduled Stage-1 or Stage-2 audits
- Post-incident: After security incidents affecting web filtering

**Quality Assurance:**
Dashboard accuracy should be validated by:
- Spot-checking consolidated data against source workbooks
- Verifying compliance calculations are correct
- Confirming gap analysis completeness
- Validating evidence linkage integrity
- Testing "Update Links" functionality

**Common Pitfalls:**
- Missing source workbooks (dashboard shows #REF! errors)
- Incorrect file naming (links broken due to name mismatch)
- Forgetting to enable "Update Links" (stale data displayed)
- Manual data edits in dashboard (overwritten on next update)
- No version control (cannot track compliance trends)
- Dashboard used as data entry (use source assessments!)

**Critical Success Factors:**
- AUTOMATE data consolidation (no manual data entry)
- VERIFY source data accuracy before dashboard generation
- MAINTAIN version control (dated file names for trend analysis)
- REVIEW dashboard regularly with stakeholders (monthly minimum)
- TRACK action items to closure (ownership and accountability)
- PRESERVE audit trail (retain historical dashboard versions)

**Compliance Scoring Methodology:**
The dashboard uses a weighted scoring approach:
- Infrastructure (A.8.23.1): 25% weight (foundational capability)
- Network Coverage (A.8.23.2): 30% weight (deployment effectiveness)
- Policy Configuration (A.8.23.3): 25% weight (control configuration)
- Monitoring & Response (A.8.23.4): 20% weight (operational effectiveness)

Weights can be adjusted based on organizational risk profile.

**"Evidence > Theater" - Systems Engineering for Real Compliance:**
This dashboard represents genuine systems engineering consolidation, not
compliance theater. It provides objective, quantitative assessment of web
filtering control effectiveness across multiple domains, enabling risk-based
decision making and audit readiness validation.

As Richard Feynman said: "The first principle is that you must not fool
yourself — and you are the easiest person to fool."

This dashboard helps you avoid fooling yourself about compliance status.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


from datetime import datetime
# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.23.5"
WORKBOOK_NAME = "Compliance Summary Dashboard"
CONTROL_ID = "A.8.23"
CONTROL_NAME = "Web Filtering"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Unicode Constants (for cross-platform compatibility)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches A.8.23 dashboard requirements
    sheets = [
        "Executive Dashboard",
        "Gap Analysis",
        "Risk Register",
        "Remediation Roadmap",
        "KPIs & Metrics",
        "Evidence Register",
        "Action Items & Follow-up",
        "Audit & Compliance Log",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "critical_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "formula_cell": {
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_green": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_yellow": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_red": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# ============================================================================
# SECTION 2: EXECUTIVE DASHBOARD
# ============================================================================

def create_executive_dashboard(ws, styles):
    """Create Executive Dashboard with external workbook links and KPIs."""
    
    # ---------- HEADER ----------
    ws.merge_cells("A1:H1")
    ws["A1"] = (
        "ISMS-IMP-A.8.23.5 – Web Filtering Compliance Summary Dashboard\n"
        "ISO/IEC 27001:2022 - Control A.8.23: Web Filtering - Executive Overview"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 50

    # ---------- DOCUMENT INFORMATION ----------
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.8.23.5"),
        ("Report Type", "Compliance Summary Dashboard"),
        ("Related Policy", "ISMS-POL-A.8.23 (Web Filtering)"),
        ("Version", "1.0"),
        ("Report Date", ""),
        ("Reporting Period", ""),
        ("Prepared By", ""),
        ("Organization", ""),
        ("Review Cycle", "Quarterly"),
        ("Last Updated", "=TODAY()"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        elif "=" in str(value):
            ws[f"B{row}"].font = Font(color="0000FF")
        row += 1

    # ---------- OVERALL COMPLIANCE STATUS ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "OVERALL WEB FILTERING COMPLIANCE STATUS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    # Compliance Scorecard
    row += 1
    scorecard_headers = ["Metric", "Score", "Target", "Status"]
    for col_idx, header in enumerate(scorecard_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    scorecard_start = row + 1
    scorecard_metrics = [
        ("Overall Compliance Rate", "=AVERAGE(C25:C28)", "≥95%", ""),
        ("Critical Gaps", "=COUNTIFS('Gap Analysis'!I8:I207,\"Critical\",'Gap Analysis'!R8:R207,\"<>Closed\")", "0", ""),
        ("High-Risk Items", "=COUNTIFS('Risk Register'!N21:N119,\">=15\",'Risk Register'!Q21:Q119,\"<>Closed\")", "≤5", ""),
        ("Remediation Progress", "=COUNTIF('Remediation Roadmap'!H37:H236,\"Completed\")/COUNTA('Remediation Roadmap'!H37:H236)*100&\"%\"", "≥80%", ""),
    ]

    row += 1
    for metric, formula, target, status in scorecard_metrics:
        ws.cell(row=row, column=1, value=metric).border = styles["border"]
        ws.cell(row=row, column=2, value=formula).font = Font(bold=True, color="0000FF")
        ws.cell(row=row, column=2).border = styles["border"]
        ws.cell(row=row, column=3, value=target).border = styles["border"]
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=4).border = styles["border"]
        row += 1

    # Add dropdown for status column
    dv_status = DataValidation(type="list", formula1='"Green,Amber,Red"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(f"D{scorecard_start}:D{row-1}")

    # ---------- DOMAIN COMPLIANCE SCORES ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "DOMAIN COMPLIANCE SCORES (from Source Workbooks)"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    domain_headers = ["#", "Assessment Domain", "Compliance Score", "Status", "Critical Gaps", "Evidence Count", "Last Updated", "Assessor"]
    for col_idx, header in enumerate(domain_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Domain rows with EXTERNAL WORKBOOK LINKS
    # This is where the magic happens - pulling data from normalized source files!
    domain_data = [
        (
            "1",
            "Filtering Infrastructure",
            "='[ISMS-IMP-A.8.23.1.xlsx]Gap_Analysis'!B5",  # Link to score cell
            "='[ISMS-IMP-A.8.23.1.xlsx]Gap_Analysis'!C5",  # Link to status
            "=COUNTIF('[ISMS-IMP-A.8.23.1.xlsx]Gap_Analysis'!F10:F109,\"Critical\")",
            "=COUNTA('[ISMS-IMP-A.8.23.1.xlsx]Evidence_Register'!A10:A209)",
            "='[ISMS-IMP-A.8.23.1.xlsx]Instructions & Legend'!B9",  # Assessment date
            "='[ISMS-IMP-A.8.23.1.xlsx]Instructions & Legend'!B10",  # Completed by
        ),
        (
            "2",
            "Network Coverage",
            "='[ISMS-IMP-A.8.23.2.xlsx]Gap_Identification'!B5",
            "='[ISMS-IMP-A.8.23.2.xlsx]Gap_Identification'!C5",
            "=COUNTIF('[ISMS-IMP-A.8.23.2.xlsx]Gap_Identification'!F10:F109,\"Critical\")",
            "=COUNTA('[ISMS-IMP-A.8.23.2.xlsx]Evidence_Register'!A10:A209)",
            "='[ISMS-IMP-A.8.23.2.xlsx]Instructions & Legend'!B9",
            "='[ISMS-IMP-A.8.23.2.xlsx]Instructions & Legend'!B10",
        ),
        (
            "3",
            "Policy Configuration",
            "='[ISMS-IMP-A.8.23.3.xlsx]Gap_Analysis'!B5",
            "='[ISMS-IMP-A.8.23.3.xlsx]Gap_Analysis'!C5",
            "=COUNTIF('[ISMS-IMP-A.8.23.3.xlsx]Gap_Analysis'!F10:F109,\"Critical\")",
            "=COUNTA('[ISMS-IMP-A.8.23.3.xlsx]Evidence_Register'!A10:A209)",
            "='[ISMS-IMP-A.8.23.3.xlsx]Instructions & Legend'!B9",
            "='[ISMS-IMP-A.8.23.3.xlsx]Instructions & Legend'!B10",
        ),
        (
            "4",
            "Monitoring & Response",
            "='[ISMS-IMP-A.8.23.4.xlsx]Gap_Analysis'!B5",
            "='[ISMS-IMP-A.8.23.4.xlsx]Gap_Analysis'!C5",
            "=COUNTIF('[ISMS-IMP-A.8.23.4.xlsx]Gap_Analysis'!F10:F109,\"Critical\")",
            "=COUNTA('[ISMS-IMP-A.8.23.4.xlsx]Evidence_Register'!A10:A209)",
            "='[ISMS-IMP-A.8.23.4.xlsx]Instructions_Legend'!B9",
            "='[ISMS-IMP-A.8.23.4.xlsx]Instructions_Legend'!B10",
        ),
    ]

    row += 1
    for domain in domain_data:
        for col_idx, value in enumerate(domain, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = styles["border"]
            if col_idx == 3:  # Compliance score
                cell.number_format = "0.0%"
                cell.fill = styles["formula_cell"]["fill"]
            elif col_idx in [4, 5, 6, 7, 8]:  # Formula columns
                cell.fill = styles["formula_cell"]["fill"]
            if "=" in str(value):
                cell.font = Font(color="0000FF", italic=True, size=9)
        row += 1

    # ---------- KEY PERFORMANCE INDICATORS ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "KEY PERFORMANCE INDICATORS (Consolidated)"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    kpi_headers = ["Category", "Metric", "Current Value", "Target", "Status", "Trend", "Notes"]
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    kpi_metrics = [
        ("Infrastructure", "Filtering Solutions Deployed", "=COUNTA('[ISMS-IMP-A.8.23.1.xlsx]Solution_Details_Template'!B8:B27)", "≥1", "", "", "From A.8.23.1"),
        ("Infrastructure", "Solutions Fully Implemented", "=COUNTIF('[ISMS-IMP-A.8.23.1.xlsx]Solution_Details_Template'!I8:I27,\"Deployed\")/COUNTA('[ISMS-IMP-A.8.23.1.xlsx]Solution_Details_Template'!B8:B27)", "100%", "", "", ""),
        ("Coverage", "Network Segments Covered", "=COUNTIF('[ISMS-IMP-A.8.23.2.xlsx]Coverage_Matrix'!D10:D109,\"\u2705 Covered\")/COUNTA('[ISMS-IMP-A.8.23.2.xlsx]Network_Segment_Inventory'!A10:A59)", "100%", "", "", "From A.8.23.2"),
        ("Coverage", "Critical Bypass Risks", "=COUNTIF('[ISMS-IMP-A.8.23.2.xlsx]Gap_Identification'!F10:F109,\"Critical\")", "0", "", "", ""),
        ("Policy", "Threat Categories Blocked", "=COUNTIF('[ISMS-IMP-A.8.23.3.xlsx]Threat_Protection'!E10:E59,\"Blocked\")/COUNTA('[ISMS-IMP-A.8.23.3.xlsx]Threat_Protection'!A10:A59)", "≥95%", "", "", "From A.8.23.3"),
        ("Policy", "Active Exceptions", "=COUNTIF('[ISMS-IMP-A.8.23.3.xlsx]Policy_Exceptions'!G10:G109,\"Active\")", "<10", "", "", ""),
        ("Monitoring", "Alert Rules Configured", "=COUNTA('[ISMS-IMP-A.8.23.4.xlsx]Alert_Configuration'!A10:A109)", "≥20", "", "", "From A.8.23.4"),
        ("Monitoring", "SLA Compliance Rate", "=COUNTIF('[ISMS-IMP-A.8.23.4.xlsx]Incident_Response'!K10:K109,\"Met\")/COUNTA('[ISMS-IMP-A.8.23.4.xlsx]Incident_Response'!K10:K109)", "≥95%", "", "", ""),
        ("Response", "Open Critical Incidents", "=COUNTIFS('[ISMS-IMP-A.8.23.4.xlsx]Incident_Response'!D10:D109,\"Critical\",'[ISMS-IMP-A.8.23.4.xlsx]Incident_Response'!I10:I109,\"Open\")", "0", "", "", ""),
        ("Response", "False Positive Rate", "=COUNTIF('[ISMS-IMP-A.8.23.4.xlsx]False_Positive_Mgmt'!H10:H109,\"Confirmed_FP\")/COUNTA('[ISMS-IMP-A.8.23.4.xlsx]Blocked_Events_Analysis'!A10:A109)", "<1%", "", "", ""),
    ]

    row += 1
    kpi_start_row = row
    for category, metric, formula, target, status, trend, notes in kpi_metrics:
        ws.cell(row=row, column=1, value=category).border = styles["border"]
        ws.cell(row=row, column=2, value=metric).border = styles["border"]
        ws.cell(row=row, column=3, value=formula).border = styles["border"]
        ws.cell(row=row, column=3).font = Font(color="0000FF", italic=True, size=9)
        ws.cell(row=row, column=3).fill = styles["formula_cell"]["fill"]
        ws.cell(row=row, column=4, value=target).border = styles["border"]
        ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=5).border = styles["border"]
        ws.cell(row=row, column=6).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=6).border = styles["border"]
        ws.cell(row=row, column=7, value=notes).border = styles["border"]
        ws.cell(row=row, column=7).font = Font(italic=True, size=9, color="666666")
        row += 1

    # Add dropdowns
    dv_status_kpi = DataValidation(type="list", formula1='"Green,Amber,Red"', allow_blank=False)
    ws.add_data_validation(dv_status_kpi)
    dv_status_kpi.add(f"E{kpi_start_row}:E{row-1}")

    dv_trend = DataValidation(type="list", formula1='"Improving,Stable,Degrading"', allow_blank=True)
    ws.add_data_validation(dv_trend)
    dv_trend.add(f"F{kpi_start_row}:F{row-1}")

    # ---------- FOOTER NOTES ----------
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "📊 DASHBOARD NOTES"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    notes = [
        "\u2022 This dashboard consolidates data from 4 assessment workbooks via external links",
        "\u2022 Blue italic formulas pull live data from normalized source files",
        "\u2022 Click 'Update Links' when opening to refresh all data",
        "\u2022 Yellow cells are for manual status/trend assessment",
        "\u2022 Place this dashboard in same folder as ISMS-IMP-A.8.23.[1-4].xlsx files",
        "\u2022 Review quarterly or after significant changes to filtering infrastructure",
    ]

    row += 1
    for note in notes:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = note
        ws[f"A{row}"].font = Font(size=9, italic=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "🎯 \"Evidence > Theater\" - This is Systems Engineering, not Cargo Cult Compliance"
    ws[f"A{row}"].font = Font(bold=True, italic=True, size=10, color="003366")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 25

    ws.freeze_panes = "A4"

# SECTION 3: GAP ANALYSIS
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create consolidated Gap Analysis sheet from all 4 assessments."""
    
    ws.merge_cells("A1:S1")
    ws["A1"] = "CONSOLIDATED GAP ANALYSIS - A.8.23 WEB FILTERING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:S2")
    ws["A2"] = "Aggregated gaps from all 4 assessment workbooks"
    ws["A2"].font = Font(bold=True, size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers
    headers = [
        "Gap ID", "Source Assessment", "Gap Category", "Gap Description", 
        "Current State", "Target State", "Impact", "Likelihood", "Risk Score",
        "Priority", "Remediation Action", "Owner", "Target Date", 
        "Status", "Progress %", "Cost Estimate", "Dependencies", 
        "Verification Method", "Status Notes"
    ]
    
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # Instructions
    row = 5
    ws.merge_cells(f"A{row}:S{row}")
    ws[f"A{row}"] = (
        "INSTRUCTIONS: Consolidate gaps from all 4 source workbooks. "
        "Each gap should reference its source assessment. "
        "Use this sheet for executive-level gap tracking across all web filtering domains."
    )
    ws[f"A{row}"].font = Font(italic=True, size=9, color="666666")
    ws[f"A{row}"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 30

    # Example gap entries (50 rows for consolidation)
    row = 6
    example_gaps = [
        ("A.8.23.1-GAP-001", "Filtering Infrastructure", "Technology", "", "", "", "High", "Medium", "=G7*H7", "High", "", "", "", "", "", "", "", "", ""),
        ("A.8.23.2-GAP-001", "Network Coverage", "Coverage", "", "", "", "Critical", "High", "=G8*H8", "Critical", "", "", "", "", "", "", "", "", ""),
        ("A.8.23.3-GAP-001", "Policy Configuration", "Policy", "", "", "", "Medium", "Low", "=G9*H9", "Medium", "", "", "", "", "", "", "", "", ""),
        ("A.8.23.4-GAP-001", "Monitoring & Response", "Monitoring", "", "", "", "High", "Medium", "=G10*H10", "High", "", "", "", "", "", "", "", "", ""),
    ]

    for gap_data in example_gaps:
        for col_idx, value in enumerate(gap_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = styles["border"]
            if value == "":
                cell.fill = styles["input_cell"]["fill"]
            elif "=" in str(value):
                cell.font = Font(color="0000FF")
        row += 1

    # Add empty rows for additional gaps (total 200 rows)
    for _ in range(46):
        for col_idx in range(1, 20):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]
        row += 1

    # Add dropdowns
    dv_priority = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dv_priority)
    dv_priority.add(f"J7:J{row-1}")

    dv_status = DataValidation(type="list", formula1='"Open,In_Progress,Resolved,Accepted,Deferred,Closed"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"N7:N{row-1}")

    # Column widths
    widths = [15, 25, 20, 40, 30, 30, 10, 12, 12, 12, 35, 20, 15, 15, 12, 15, 25, 25, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 4: RISK REGISTER
# ============================================================================

def create_risk_register(ws, styles):
    """Create consolidated Risk Register."""
    
    ws.merge_cells("A1:Q1")
    ws["A1"] = "RISK REGISTER - A.8.23 WEB FILTERING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Risk matrix reference
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "RISK SCORING MATRIX"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    row += 1
    ws[f"A{row}"] = "Impact / Likelihood"
    ws[f"B{row}"] = "Low (1)"
    ws[f"C{row}"] = "Medium (2)"
    ws[f"D{row}"] = "High (3)"
    ws[f"E{row}"] = "Critical (4)"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].font = Font(bold=True, size=9)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"{col}{row}"].border = styles["border"]

    risk_matrix = [
        ("Critical (4)", "4", "8", "12", "16"),
        ("High (3)", "3", "6", "9", "12"),
        ("Medium (2)", "2", "4", "6", "8"),
        ("Low (1)", "1", "2", "3", "4"),
    ]

    row += 1
    for risk_row in risk_matrix:
        ws[f"A{row}"] = risk_row[0]
        ws[f"A{row}"].font = Font(bold=True, size=9)
        ws[f"A{row}"].border = styles["border"]
        for col_idx, val in enumerate(risk_row[1:], start=2):
            cell = ws.cell(row=row, column=col_idx, value=int(val))
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="center")
            # Color code by score
            score = int(val)
            if score >= 12:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif score >= 6:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row += 1

    # Risk register table
    row += 2
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "IDENTIFIED RISKS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    headers = [
        "Risk ID", "Source", "Risk Category", "Risk Description", 
        "Threat Source", "Vulnerability", "Impact", "Likelihood", 
        "Inherent Risk", "Controls", "Control Effectiveness", 
        "Residual Impact", "Residual Likelihood", "Residual Risk", 
        "Treatment", "Owner", "Status"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # Add 100 risk entry rows
    row += 1
    for i in range(100):
        for col_idx in range(1, 18):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]
            
            # Add formulas for risk calculations
            if col_idx == 9:  # Inherent Risk
                cell.value = f"=G{row}*H{row}"
                cell.font = Font(color="0000FF")
            elif col_idx == 14:  # Residual Risk
                cell.value = f"=L{row}*M{row}"
                cell.font = Font(color="0000FF")
        row += 1

    # Add dropdowns
    # FIX: Split comma-separated ranges into separate add() calls
    dv_impact = DataValidation(type="list", formula1='"1,2,3,4"', allow_blank=True)
    ws.add_data_validation(dv_impact)
    dv_impact.add(f"G{row-100}:G{row-1}")  # Inherent Impact
    dv_impact.add(f"L{row-100}:L{row-1}")  # Residual Impact

    # FIX: Split comma-separated ranges into separate add() calls
    dv_likelihood = DataValidation(type="list", formula1='"1,2,3,4"', allow_blank=True)
    ws.add_data_validation(dv_likelihood)
    dv_likelihood.add(f"H{row-100}:H{row-1}")  # Inherent Likelihood
    dv_likelihood.add(f"M{row-100}:M{row-1}")  # Residual Likelihood

    dv_treatment = DataValidation(type="list", formula1='"Accept,Mitigate,Transfer,Avoid"', allow_blank=True)
    ws.add_data_validation(dv_treatment)
    dv_treatment.add(f"O{row-100}:O{row-1}")

    dv_status = DataValidation(type="list", formula1='"Open,Monitoring,Resolved,Closed"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"Q{row-100}:Q{row-1}")

    # Column widths
    widths = [15, 20, 20, 40, 25, 25, 10, 12, 12, 30, 15, 10, 12, 12, 15, 20, 15]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A20"


# ============================================================================
# SECTION 5: REMEDIATION ROADMAP
# ============================================================================

def create_remediation_roadmap(ws, styles):
    """Create Remediation Roadmap with timeline tracking."""
    
    ws.merge_cells("A1:P1")
    ws["A1"] = "REMEDIATION ROADMAP - A.8.23 WEB FILTERING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:P2")
    ws["A2"] = "Action plan for closing identified gaps and reducing risks"
    ws["A2"].font = Font(bold=True, size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Roadmap summary
    row = 4
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "REMEDIATION SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    summary_metrics = [
        ("Total Action Items", "=COUNTA(A37:A236)"),
        ("Completed", "=COUNTIF(H37:H236,\"Completed\")"),
        ("In Progress", "=COUNTIF(H37:H236,\"In Progress\")"),
        ("Not Started", "=COUNTIF(H37:H236,\"Not Started\")"),
        ("Overdue", "=COUNTIFS(H37:H236,\"<>Completed\",E37:E236,\"<\"&TODAY())"),
        ("Overall Progress", "=COUNTIF(H37:H236,\"Completed\")/COUNTA(A37:A236)"),
    ]

    row += 1
    for label, formula in summary_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(color="0000FF", bold=True)
        ws[f"B{row}"].border = styles["border"]
        if "Progress" in label:
            ws[f"B{row}"].number_format = "0.0%"
        row += 1

    # Timeline phases
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "IMPLEMENTATION PHASES"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    phases = [
        ("Phase 1: Critical (0-30 days)", "Immediate action required", "Critical gaps, security risks"),
        ("Phase 2: High (30-90 days)", "Quick wins, high impact", "Infrastructure, coverage gaps"),
        ("Phase 3: Medium (90-180 days)", "Process improvements", "Policy, monitoring enhancements"),
        ("Phase 4: Low (180-365 days)", "Optimization", "Fine-tuning, documentation"),
    ]

    row += 1
    for phase, description, focus in phases:
        ws[f"A{row}"] = phase
        ws[f"A{row}"].font = Font(bold=True, size=10)
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = description
        ws[f"B{row}"].border = styles["border"]
        ws.merge_cells(f"C{row}:F{row}")
        ws[f"C{row}"] = focus
        ws[f"C{row}"].border = styles["border"]
        ws[f"C{row}"].font = Font(italic=True, size=9)
        row += 1

    # Action item table
    row += 2
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "ACTION ITEMS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    headers = [
        "Action ID", "Phase", "Related Gap/Risk", "Action Description",
        "Target Date", "Assigned Owner", "Department", "Status",
        "Progress %", "Actual Completion", "Budget", "Dependencies",
        "Success Criteria", "Verification", "Notes", "Last Updated"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # Add 200 action item rows
    row += 1
    for i in range(200):
        for col_idx in range(1, 17):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]
        row += 1

    # Add dropdowns
    dv_phase = DataValidation(type="list", formula1='"Phase 1,Phase 2,Phase 3,Phase 4"', allow_blank=True)
    ws.add_data_validation(dv_phase)
    dv_phase.add(f"B37:B236")

    dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Completed,Blocked,Cancelled"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"H37:H236")

    # Column widths
    widths = [15, 15, 20, 40, 15, 20, 15, 15, 12, 15, 15, 25, 30, 20, 30, 15]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A37"

# SECTION 6: KPIs & METRICS
# ============================================================================

def create_kpis_metrics(ws, styles):
    """Create KPIs & Metrics tracking sheet."""
    
    ws.merge_cells("A1:M1")
    ws["A1"] = "KPIs & METRICS - A.8.23 WEB FILTERING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # KPI Categories
    categories = [
        {
            "name": "INFRASTRUCTURE METRICS",
            "kpis": [
                ("Total Filtering Solutions", "Count", "", "≥1", ""),
                ("Deployment Coverage", "%", "", "100%", ""),
                ("Solution Availability", "%", "", "≥99%", ""),
                ("Performance - Latency", "ms", "", "≤100ms", ""),
                ("Licensing Compliance", "%", "", "100%", ""),
            ]
        },
        {
            "name": "COVERAGE METRICS",
            "kpis": [
                ("Network Segments Covered", "%", "", "100%", ""),
                ("Users Protected", "Count", "", "All", ""),
                ("Devices Protected", "Count", "", "All", ""),
                ("Bypass Vulnerabilities", "Count", "", "0", ""),
                ("Coverage Verification Rate", "%", "", "100%", ""),
            ]
        },
        {
            "name": "POLICY METRICS",
            "kpis": [
                ("Threat Categories Blocked", "%", "", "≥95%", ""),
                ("Content Categories Configured", "Count", "", "≥30", ""),
                ("Active Policy Exceptions", "Count", "", "≤10", ""),
                ("Exception Review Compliance", "%", "", "100%", ""),
                ("Policy Update Frequency", "Days", "", "≤90", ""),
            ]
        },
        {
            "name": "MONITORING METRICS",
            "kpis": [
                ("Log Collection Rate", "%", "", "100%", ""),
                ("Alert Rules Active", "Count", "", "≥20", ""),
                ("Alert Response Time", "Minutes", "", "≤30", ""),
                ("Dashboard Availability", "%", "", "≥99%", ""),
                ("SIEM Integration Health", "%", "", "100%", ""),
            ]
        },
        {
            "name": "INCIDENT METRICS",
            "kpis": [
                ("Open Critical Incidents", "Count", "", "0", ""),
                ("Mean Time to Detect (MTTD)", "Minutes", "", "≤15", ""),
                ("Mean Time to Respond (MTTR)", "Hours", "", "≤4", ""),
                ("SLA Compliance Rate", "%", "", "≥95%", ""),
                ("False Positive Rate", "%", "", "≤1%", ""),
            ]
        },
        {
            "name": "COMPLIANCE METRICS",
            "kpis": [
                ("Overall Compliance Score", "%", "", "≥95%", ""),
                ("Critical Gaps Open", "Count", "", "0", ""),
                ("Evidence Collection Rate", "%", "", "100%", ""),
                ("Audit Readiness", "Status", "", "Ready", ""),
                ("Policy Review Currency", "Days", "", "≤90", ""),
            ]
        },
    ]

    row = 3
    for category in categories:
        # Category header
        ws.merge_cells(f"A{row}:M{row}")
        ws[f"A{row}"] = category["name"]
        ws[f"A{row}"].font = styles["section_header"]["font"]
        ws[f"A{row}"].fill = styles["section_header"]["fill"]
        ws[f"A{row}"].alignment = styles["section_header"]["alignment"]
        
        row += 1
        # Column headers
        headers = ["KPI Name", "Unit", "Current Value", "Target", "Status", "Trend", 
                   "Last Month", "YTD Average", "Data Source", "Update Frequency", 
                   "Owner", "Collection Method", "Notes"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = styles["column_header"]["font"]
            cell.fill = styles["column_header"]["fill"]
            cell.alignment = styles["column_header"]["alignment"]
            cell.border = styles["border"]
        
        row += 1
        # KPI rows
        for kpi_name, unit, current, target, status in category["kpis"]:
            ws[f"A{row}"] = kpi_name
            ws[f"A{row}"].border = styles["border"]
            ws[f"B{row}"] = unit
            ws[f"B{row}"].border = styles["border"]
            ws[f"C{row}"].fill = styles["input_cell"]["fill"]
            ws[f"C{row}"].border = styles["border"]
            ws[f"D{row}"] = target
            ws[f"D{row}"].border = styles["border"]
            for col in ["E", "F", "G", "H", "I", "J", "K", "L", "M"]:
                ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
                ws[f"{col}{row}"].border = styles["border"]
            row += 1
        
        row += 1  # Blank row between categories

    # Add dropdowns
    dv_status = DataValidation(type="list", formula1='"Green,Amber,Red"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("E4:E100")

    dv_trend = DataValidation(type="list", formula1='"Improving,Stable,Degrading,New"', allow_blank=True)
    ws.add_data_validation(dv_trend)
    dv_trend.add("F4:F100")

    dv_frequency = DataValidation(type="list", formula1='"Real-Time,Hourly,Daily,Weekly,Monthly"', allow_blank=True)
    ws.add_data_validation(dv_frequency)
    dv_frequency.add("J4:J100")

    # Column widths
    widths = [30, 10, 15, 15, 12, 12, 15, 15, 25, 18, 20, 20, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create consolidated Evidence Register."""
    
    ws.merge_cells("A1:O1")
    ws["A1"] = "EVIDENCE REGISTER - A.8.23 WEB FILTERING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:O2")
    ws["A2"] = "Centralized evidence tracking for audit and compliance verification"
    ws["A2"].font = Font(bold=True, size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Evidence summary
    row = 4
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "EVIDENCE SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    summary_metrics = [
        ("Total Evidence Items", "=COUNTA(A20:A519)"),
        ("Verified", "=COUNTIF(L21:L519,\"Verified\")"),
        ("Pending Verification", "=COUNTIF(L21:L519,\"Pending\")"),
        ("Expired/Missing", "=COUNTIF(L21:L519,\"Expired\")"),
        ("Verification Rate", "=COUNTIF(L21:L519,\"Verified\")/COUNTA(A20:A519)"),
    ]

    row += 1
    for label, formula in summary_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(color="0000FF", bold=True)
        ws[f"B{row}"].border = styles["border"]
        if "Rate" in label:
            ws[f"B{row}"].number_format = "0.0%"
        row += 1

    # Evidence retention by category
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "EVIDENCE BY CATEGORY"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    ws[f"A{row}"] = "Category"
    ws[f"B{row}"] = "Count"
    ws[f"C{row}"] = "Verified"
    ws[f"D{row}"] = "Verification %"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws[f"{col}{row}"].border = styles["border"]

    categories = [
        "Infrastructure", "Coverage", "Policy", "Monitoring",
        "Incident Response", "Compliance", "Audit"
    ]

    row += 1
    for cat in categories:
        ws[f"A{row}"] = cat
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = f'=COUNTIF(D21:D519,"{cat}")'
        ws[f"B{row}"].font = Font(color="0000FF")
        ws[f"B{row}"].border = styles["border"]
        ws[f"C{row}"] = f'=COUNTIFS(D21:D519,"{cat}",L21:L519,"Verified")'
        ws[f"C{row}"].font = Font(color="0000FF")
        ws[f"C{row}"].border = styles["border"]
        ws[f"D{row}"] = f"=C{row}/B{row}"
        ws[f"D{row}"].font = Font(color="0000FF")
        ws[f"D{row}"].border = styles["border"]
        ws[f"D{row}"].number_format = "0.0%"
        row += 1

    # Evidence table
    row += 1
    ws.merge_cells(f"A{row}:O{row}")
    ws[f"A{row}"] = "EVIDENCE ITEMS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    headers = [
        "Evidence ID", "Source Assessment", "Related Requirement", "Category",
        "Evidence Type", "Description", "File Location", "Collection Date",
        "Collected By", "Retention Period", "Expiry Date", "Verification Status",
        "Verified By", "Verification Date", "Notes"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # Add 500 evidence entry rows
    row += 1
    for i in range(500):
        for col_idx in range(1, 16):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]
            
            # Add formula for expiry date
            if col_idx == 11:  # Expiry Date
                cell.value = f"=H{row}+J{row}"
                cell.font = Font(color="0000FF")
        row += 1

    # Add dropdowns
    dv_category = DataValidation(type="list", formula1='"Infrastructure,Coverage,Policy,Monitoring,Incident Response,Compliance,Audit"', allow_blank=True)
    ws.add_data_validation(dv_category)
    dv_category.add("D21:D519")

    dv_type = DataValidation(type="list", formula1='"Screenshot,Configuration_Export,Log_Sample,Report,Procedure,Email,Audit_Report,Test_Results,Policy_Document,Meeting_Minutes,Other"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add("E20:E519")

    dv_status = DataValidation(type="list", formula1='"Verified,Pending,Not_Verified,Expired"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("L21:L519")

    # Column widths
    widths = [15, 20, 25, 18, 20, 40, 30, 15, 20, 15, 15, 18, 20, 15, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A20"


# ============================================================================
# SECTION 8: ACTION ITEMS & FOLLOW-UP
# ============================================================================

def create_action_items(ws, styles):
    """Create Action Items & Follow-up tracking sheet."""
    
    ws.merge_cells("A1:N1")
    ws["A1"] = "ACTION ITEMS & FOLLOW-UP - A.8.23 WEB FILTERING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Action summary
    row = 3
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "ACTION SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    summary_metrics = [
        ("Total Actions", "=COUNTA(A20:A219)"),
        ("Overdue", "=COUNTIFS(F20:F219,\"<\"&TODAY(),G20:G219,\"<>Completed\")"),
        ("Due This Week", "=COUNTIFS(F20:F219,\">=\"&TODAY(),F20:F219,\"<=\"&TODAY()+7,G20:G219,\"<>Completed\")"),
        ("Completed This Month", "=COUNTIFS(G20:G219,\"Completed\",H20:H219,\">=\"&EOMONTH(TODAY(),-1)+1)"),
        ("Completion Rate", "=COUNTIF(G20:G219,\"Completed\")/COUNTA(A20:A219)"),
    ]

    row += 1
    for label, formula in summary_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(color="0000FF", bold=True)
        ws[f"B{row}"].border = styles["border"]
        if "Rate" in label:
            ws[f"B{row}"].number_format = "0.0%"
        row += 1

    # Actions by priority
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ACTIONS BY PRIORITY"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    ws[f"A{row}"] = "Priority"
    ws[f"B{row}"] = "Total"
    ws[f"C{row}"] = "Open"
    ws[f"D{row}"] = "Overdue"
    ws[f"E{row}"] = "Completed"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws[f"{col}{row}"].border = styles["border"]

    priorities = ["Critical", "High", "Medium", "Low"]
    row += 1
    for priority in priorities:
        ws[f"A{row}"] = priority
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = f'=COUNTIF(E20:E219,"{priority}")'
        ws[f"B{row}"].font = Font(color="0000FF")
        ws[f"B{row}"].border = styles["border"]
        ws[f"C{row}"] = f'=COUNTIFS(E20:E219,"{priority}",G20:G219,"<>Completed")'
        ws[f"C{row}"].font = Font(color="0000FF")
        ws[f"C{row}"].border = styles["border"]
        ws[f"D{row}"] = f'=COUNTIFS(E20:E219,"{priority}",F20:F219,"<"&TODAY(),G20:G219,"<>Completed")'
        ws[f"D{row}"].font = Font(color="0000FF")
        ws[f"D{row}"].border = styles["border"]
        ws[f"E{row}"] = f'=COUNTIFS(E20:E219,"{priority}",G20:G219,"Completed")'
        ws[f"E{row}"].font = Font(color="0000FF")
        ws[f"E{row}"].border = styles["border"]
        row += 1

    # Action items table
    row += 1
    ws.merge_cells(f"A{row}:N{row}")
    ws[f"A{row}"] = "ACTION ITEMS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    headers = [
        "Action ID", "Source", "Related Gap/Risk", "Action Description",
        "Priority", "Due Date", "Status", "Completion Date",
        "Assigned To", "Department", "Progress Notes", "Blocker",
        "Next Steps", "Last Updated"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # Add 200 action item rows
    row += 1
    for i in range(200):
        for col_idx in range(1, 15):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]
        row += 1

    # Add dropdowns
    dv_priority = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dv_priority)
    dv_priority.add("E20:E219")

    dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Blocked,Completed,Cancelled"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("G20:G219")

    dv_blocker = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_blocker)
    dv_blocker.add("L20:L219")

    # Column widths
    widths = [15, 20, 20, 40, 12, 15, 15, 15, 20, 20, 35, 10, 30, 15]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A20"

# SECTION 9: AUDIT & COMPLIANCE LOG
# ============================================================================

def create_audit_log(ws, styles):
    """Create Audit & Compliance Log sheet."""
    
    ws.merge_cells("A1:M1")
    ws["A1"] = "AUDIT & COMPLIANCE LOG - A.8.23 WEB FILTERING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:M2")
    ws["A2"] = "Tracking audit activities, findings, and compliance verification"
    ws["A2"].font = Font(bold=True, size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Audit summary
    row = 4
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "AUDIT SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    summary_fields = [
        ("Last Internal Audit", ""),
        ("Last External Audit", ""),
        ("Next Scheduled Audit", ""),
        ("Total Findings (All Time)", "=COUNTA(A20:A119)"),
        ("Open Findings", "=COUNTIF(I21:I119,\"Open\")"),
        ("Closed Findings", "=COUNTIF(I21:I119,\"Closed\")"),
        ("Average Resolution Time (Days)", "=AVERAGE(J20:J119)"),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = styles["border"]
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        else:
            ws[f"B{row}"] = value
            ws[f"B{row}"].font = Font(color="0000FF", bold=True)
            ws[f"B{row}"].border = styles["border"]
        row += 1

    # Findings by severity
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "FINDINGS BY SEVERITY"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    ws[f"A{row}"] = "Severity"
    ws[f"B{row}"] = "Total"
    ws[f"C{row}"] = "Open"
    ws[f"D{row}"] = "Closed"
    ws[f"E{row}"] = "Closure Rate"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws[f"{col}{row}"].border = styles["border"]

    severities = ["Critical", "High", "Medium", "Low", "Observation"]
    row += 1
    for severity in severities:
        ws[f"A{row}"] = severity
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = f'=COUNTIF(G21:G119,"{severity}")'
        ws[f"B{row}"].font = Font(color="0000FF")
        ws[f"B{row}"].border = styles["border"]
        ws[f"C{row}"] = f'=COUNTIFS(G21:G119,"{severity}",I21:I119,"Open")'
        ws[f"C{row}"].font = Font(color="0000FF")
        ws[f"C{row}"].border = styles["border"]
        ws[f"D{row}"] = f'=COUNTIFS(G21:G119,"{severity}",I21:I119,"Closed")'
        ws[f"D{row}"].font = Font(color="0000FF")
        ws[f"D{row}"].border = styles["border"]
        ws[f"E{row}"] = f"=D{row}/B{row}"
        ws[f"E{row}"].font = Font(color="0000FF")
        ws[f"E{row}"].border = styles["border"]
        ws[f"E{row}"].number_format = "0.0%"
        row += 1

    # Audit log table
    row += 1
    ws.merge_cells(f"A{row}:M{row}")
    ws[f"A{row}"] = "AUDIT FINDINGS & COMPLIANCE RECORDS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    row += 1
    headers = [
        "Finding ID", "Audit Date", "Audit Type", "Auditor/Source",
        "Finding Category", "Finding Description", "Severity",
        "Affected Area", "Status", "Resolution Days", "Resolution Date",
        "Remediation Action", "Notes"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # Add 100 audit record rows
    row += 1
    for i in range(100):
        for col_idx in range(1, 14):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]
            
            # Add formula for resolution days
            if col_idx == 10:  # Resolution Days
                cell.value = f"=IF(K{row}=\"\",\"\",K{row}-B{row})"
                cell.font = Font(color="0000FF")
        row += 1

    # Add dropdowns
    dv_audit_type = DataValidation(type="list", formula1='"Internal,External,Certification,Surveillance,Self-Assessment"', allow_blank=True)
    ws.add_data_validation(dv_audit_type)
    dv_audit_type.add("C20:C119")

    dv_category = DataValidation(type="list", formula1='"Infrastructure,Coverage,Policy,Monitoring,Documentation,Process,Technical,Organizational"', allow_blank=True)
    ws.add_data_validation(dv_category)
    dv_category.add("E20:E119")

    dv_severity = DataValidation(type="list", formula1='"Critical,High,Medium,Low,Observation"', allow_blank=True)
    ws.add_data_validation(dv_severity)
    dv_severity.add("G21:G119")

    dv_status = DataValidation(type="list", formula1='"Open,In Progress,Pending Verification,Closed,Accepted Risk"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("I21:I119")

    # Column widths
    widths = [15, 15, 20, 25, 20, 40, 15, 25, 18, 15, 15, 35, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A20"


# ============================================================================
# SECTION 10: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create Approval Sign-Off sheet."""
    
    ws.merge_cells("A1:E1")
    ws["A1"] = "APPROVAL SIGN-OFF - A.8.23 WEB FILTERING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    ws["A2"] = "Multi-level approval workflow for compliance dashboard"
    ws["A2"].font = Font(bold=True, size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Assessment Summary
    row = 4
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY FOR APPROVAL"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    ws[f"A{row}"].alignment = styles["section_header"]["alignment"]

    summary_items = [
        ("Overall Compliance Score", "='Executive Dashboard'!B17", ""),
        ("Total Critical Gaps", "='Executive Dashboard'!B18", ""),
        ("Total High-Risk Items", "='Executive Dashboard'!B19", ""),
        ("Evidence Items Collected", "='Evidence Register'!B5", ""),
        ("Open Action Items", "='Action Items & Follow-up'!B6", ""),
        ("Audit Readiness Status", "", "Ready/Conditional/Not Ready"),
    ]

    row += 1
    for label, formula, note in summary_items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = styles["border"]
        if formula:
            ws[f"B{row}"] = formula
            ws[f"B{row}"].font = Font(color="0000FF", bold=True)
            ws[f"B{row}"].border = styles["border"]
        else:
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        if note:
            ws.merge_cells(f"C{row}:E{row}")
            ws[f"C{row}"] = note
            ws[f"C{row}"].font = Font(italic=True, size=9, color="666666")
        row += 1

    # Assessment Status
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT STATUS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    row += 1
    ws[f"A{row}"] = "Assessment Status:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]

    # Add dropdown for assessment status
    status_cell_row = row
    dv_status = DataValidation(
        type="list",
        formula1='"Draft,Under Review,Approved,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{status_cell_row}"])

    # Completed By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    completion_fields = ["Name", "Role/Title", "Department", "Email", "Date"]
    row += 1
    for field in completion_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Reviewed By (Information Security Officer)
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (INFORMATION SECURITY OFFICER)"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    review_fields = ["Name", "Date", "Review Notes"]
    row += 1
    for field in review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Approved By (CISO)
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    approval_fields = ["Name", "Date", "Approval Decision", "Conditions/Notes"]
    row += 1
    for field in approval_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Approval Decision dropdown (3rd field from approval section start)
    decision_cell_row = row - 2
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{decision_cell_row}"])

    # Next Review Details
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    next_review_fields = ["Next Review Date", "Review Responsible", "Special Considerations"]
    row += 1
    for field in next_review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # CISO Certification Statement
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "CISO CERTIFICATION STATEMENT"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    certification = (
        "I certify that this consolidated assessment of ISO/IEC 27001:2022 Control A.8.23 "
        "(Web Filtering) has been reviewed and represents an accurate view of our organization's "
        "compliance status across all four assessment domains (Infrastructure, Coverage, Policy, "
        "Monitoring & Response). Identified gaps have remediation plans with assigned owners and "
        "target dates. This dashboard consolidates data from completed assessments and is suitable "
        "for presentation to internal and external auditors."
    )
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = certification
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws[f"A{row}"].border = styles["border"]
    ws.row_dimensions[row].height = 90

    # Feynman Quote
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "🎯 \"The first principle is that you must not fool yourself – and you are the easiest person to fool.\" – Richard Feynman"
    ws[f"A{row}"].font = Font(bold=True, italic=True, size=10, color="003366")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", wrap_text=True)

    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "This is Systems Engineering. This is Evidence > Theater."
    ws[f"A{row}"].font = Font(bold=True, size=9, color="666666")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 11: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.8.23.5 - Compliance Summary Dashboard Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.23: Web Filtering")
    logger.info("=" * 80)
    logger.info("\n🎯 Systems Engineering Approach - Consolidated Oversight Dashboard")
    logger.info("   This dashboard pulls live data from 4 assessment workbooks\n")

    wb = create_workbook()
    styles = setup_styles()

    logger.info("\n[1/9] Creating Executive Dashboard sheet...")
    create_executive_dashboard(wb["Executive Dashboard"], styles)

    logger.info("[2/9] Creating Gap Analysis sheet...")
    create_gap_analysis(wb["Gap Analysis"], styles)

    logger.info("[3/9] Creating Risk Register sheet...")
    create_risk_register(wb["Risk Register"], styles)

    logger.info("[4/9] Creating Remediation Roadmap sheet...")
    create_remediation_roadmap(wb["Remediation Roadmap"], styles)

    logger.info("[5/9] Creating KPIs & Metrics sheet...")
    create_kpis_metrics(wb["KPIs & Metrics"], styles)

    logger.info("[6/9] Creating Evidence Register sheet...")
    create_evidence_register(wb["Evidence Register"], styles)

    logger.info("[7/9] Creating Action Items & Follow-up sheet...")
    create_action_items(wb["Action Items & Follow-up"], styles)

    logger.info("[8/9] Creating Audit & Compliance Log sheet...")
    create_audit_log(wb["Audit & Compliance Log"], styles)

    logger.info("[9/9] Creating Approval Sign-Off sheet...")
    create_approval_signoff(wb["Approval Sign-Off"], styles)

    filename = f"ISMS-IMP-A.8.23.5_Compliance_Summary_Dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    logger.info(f"\n\u2705 SUCCESS: {filename}")
    logger.info("\n" + "=" * 80)
    logger.info("COMPLIANCE SUMMARY DASHBOARD - COMPLETE")
    logger.info("=" * 80)
    logger.info("\nWorkbook Structure (9 Sheets):")
    logger.info("  1. Executive Dashboard - Overall compliance with external links to source workbooks")
    logger.info("  2. Gap Analysis - 200 gap entries consolidated from all 4 assessments")
    logger.info("  3. Risk Register - 100 risk entries with inherent/residual scoring")
    logger.info("  4. Remediation Roadmap - 200 remediation items with timeline tracking")
    logger.info("  5. KPIs & Metrics - 50+ KPIs across 6 categories")
    logger.info("  6. Evidence Register - 500 evidence entries with retention tracking")
    logger.info("  7. Action Items & Follow-up - 200 action items with status tracking")
    logger.info("  8. Audit & Compliance Log - 100 audit records")
    logger.info("  9. Approval Sign-Off - Multi-level approval workflow")
    
    logger.info("\nExternal Links Configuration:")
    logger.info("  \u2022 Dashboard references 4 normalized source workbooks:")
    logger.info("    - ISMS-IMP-A.8.23.1.xlsx (Filtering Infrastructure)")
    logger.info("    - ISMS-IMP-A.8.23.2.xlsx (Network Coverage)")
    logger.info("    - ISMS-IMP-A.8.23.3.xlsx (Policy Configuration)")
    logger.info("    - ISMS-IMP-A.8.23.4.xlsx (Monitoring & Response)")
    
    logger.info("\nIMPORTANT SETUP STEPS:")
    logger.info("  1. Run normalization script FIRST:")
    logger.info("     python3 normalize_assessment_files_a823.py")
    logger.info("")
    logger.info("  2. Place this dashboard in the Dashboard_Sources folder")
    logger.info("     alongside the 4 normalized source workbooks")
    logger.info("")
    logger.info("  3. Open dashboard and click 'Update Links' when prompted")
    logger.info("")
    logger.info("  4. Executive Dashboard will auto-populate with compliance data!")
    logger.info("")
    logger.info("  5. Complete manual entry sections:")
    logger.info("     - Gap Analysis (consolidated gaps)")
    logger.info("     - Risk Register (identified risks)")
    logger.info("     - Remediation Roadmap (action plans)")
    logger.info("     - KPIs (current metric values)")
    logger.info("     - Evidence Register (supporting documentation)")
    logger.info("")
    logger.info("=" * 80)
    logger.info("\n🎄 ALL 5 WEB FILTERING ASSESSMENT TOOLS COMPLETE! 🎄")
    logger.info("\nYou now have:")
    logger.info("  \u2705 4 Assessment workbooks (Infrastructure, Coverage, Policy, Monitoring)")
    logger.info("  \u2705 1 Normalization script (prepares files for dashboard)")
    logger.info("  \u2705 1 Consolidated dashboard (executive oversight)")
    logger.info("\n🎯 Evidence > Theater - This is Systems Engineering, not Cargo Cult!")
    logger.info("=" * 80 + "\n")


if __name__ == "__main__":
    main()


# ============================================================================
# END OF SCRIPT - COMPLETE COMPLIANCE SUMMARY DASHBOARD GENERATOR
# ============================================================================

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
