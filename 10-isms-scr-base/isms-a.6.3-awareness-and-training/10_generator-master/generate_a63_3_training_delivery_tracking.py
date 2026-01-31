#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.3.3 - Training Delivery and Tracking Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.3: Information Security Awareness, Education and Training
Assessment Domain 3 of 4: Training Delivery and Tracking

Reference Pattern: Based on ISMS-IMP-A.6.3.3 specification

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates an Excel workbook for tracking training delivery,
completion status, assessment results, and remediation activities.

**Purpose:**
Track training completion, measure assessment results, manage remediation,
and verify competency across all personnel.

**Generated Workbook Structure:**
1. Instructions - Operational guidance
2. Personnel_Register - Master personnel list
3. Completion_Tracking - Training completion records
4. Assessment_Results - Test scores and competency
5. Simulation_Results - Phishing and behavioral testing
6. Remediation_Tracking - Failed assessments and corrective actions
7. Compliance_Summary - Aggregated metrics by department/tier
8. Overdue_Alerts - Real-time overdue training alerts
9. Evidence_Register - Audit evidence
10. Dashboard - Visual summary
11. Approval_Sign_Off - Monthly/quarterly attestation

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.3.3"
WORKBOOK_NAME = "Training Delivery Tracking"
CONTROL_ID = "A.6.3"
CONTROL_NAME = "Information Security Awareness, Education and Training"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "overdue": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "completed": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "warning": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
    }


# =============================================================================
# WORKBOOK CREATION
# =============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Personnel_Register",
        "Completion_Tracking",
        "Assessment_Results",
        "Simulation_Results",
        "Remediation_Tracking",
        "Compliance_Summary",
        "Overdue_Alerts",
        "Evidence_Register",
        "Dashboard",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Training Completion Tracking and Effectiveness Measurement"),
        ("Related Policy", "ISMS-POL-A.6.3, Sections 2.5-2.6"),
        ("Version", "1.0"),
        ("Reporting Period", ""),
        ("Completed By", ""),
        ("Review Cycle", "Continuous tracking + Monthly reporting"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Maintain Personnel_Register with current employee data (sync with HRIS).",
        "2. Record training completions in Completion_Tracking (or import from LMS).",
        "3. Track assessment results in Assessment_Results sheet.",
        "4. Import phishing simulation results to Simulation_Results.",
        "5. Track remediation cases in Remediation_Tracking.",
        "6. Review Compliance_Summary for department-level metrics.",
        "7. Monitor Overdue_Alerts for proactive follow-up.",
        "8. Generate monthly reports from Dashboard.",
        "9. Complete Approval_Sign_Off for attestation.",
    ]

    row += 1
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    row += 1
    ws[f"A{row}"] = "STATUS LEGEND"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    legend = [
        ("Completed", "Training completed on time", "C6EFCE"),
        ("In Progress", "Training assigned, not yet completed", "FFFFCC"),
        ("Overdue", "Past due date, not completed", "FFC7CE"),
        ("Not Started", "Assigned but not begun", "D9D9D9"),
    ]

    row += 1
    for status, description, color in legend:
        ws[f"A{row}"] = status
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = description
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


def create_personnel_register_sheet(ws, styles):
    """Create Personnel_Register sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "PERSONNEL REGISTER\nMaster list of all personnel requiring training"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Employee_ID", 15),
        ("Full_Name", 25),
        ("Department", 20),
        ("Role_Title", 25),
        ("Training_Tier", 15),
        ("Employment_Type", 18),
        ("Start_Date", 12),
        ("Status", 12),
        ("Manager", 25),
        ("Email", 30),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Dropdowns
    dv_tier = DataValidation(type="list", formula1='"Tier 1,Tier 2,Tier 3,Tier 4,Tier 5,Tier 6,Tier 7"', allow_blank=False)
    dv_emp_type = DataValidation(type="list", formula1='"Full-Time,Part-Time,Contractor,Consultant,Intern"', allow_blank=False)
    dv_status = DataValidation(type="list", formula1='"Active,On Leave,Terminated"', allow_blank=False)

    ws.add_data_validation(dv_tier)
    ws.add_data_validation(dv_emp_type)
    ws.add_data_validation(dv_status)

    for r in range(4, 504):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_tier.add(ws.cell(row=r, column=5))
        dv_emp_type.add(ws.cell(row=r, column=6))
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"


def create_completion_tracking_sheet(ws, styles):
    """Create Completion_Tracking sheet."""
    ws.merge_cells("A1:O1")
    ws["A1"] = "COMPLETION TRACKING\nRecord of all training completions"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Record_ID", 12),
        ("Employee_ID", 12),
        ("Employee_Name", 25),
        ("Module_ID", 15),
        ("Module_Title", 35),
        ("Assigned_Date", 12),
        ("Due_Date", 12),
        ("Completion_Date", 15),
        ("Status", 15),
        ("Days_Overdue", 12),
        ("Assessment_Score", 15),
        ("Pass_Fail", 10),
        ("Attempts", 10),
        ("Certificate_ID", 20),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Dropdowns
    dv_status = DataValidation(type="list", formula1='"Completed,In Progress,Overdue,Not Started"', allow_blank=False)
    dv_pass = DataValidation(type="list", formula1='"Pass,Fail,N/A"', allow_blank=False)

    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_pass)

    for r in range(4, 1004):
        ws.cell(row=r, column=1, value=f"TRK-{r-3:04d}").font = Font(color="808080")

        for c in range(2, 16):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Status formula
        ws.cell(row=r, column=9, value=f'=IF(H{r}<>"","Completed",IF(TODAY()>G{r},"Overdue",IF(F{r}<>"","In Progress","Not Started")))')

        # Days Overdue formula
        ws.cell(row=r, column=10, value=f'=IF(I{r}="Overdue",TODAY()-G{r},0)')
        ws.cell(row=r, column=10).font = Font(bold=True, color="C00000")

        dv_pass.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "A4"


def create_assessment_results_sheet(ws, styles):
    """Create Assessment_Results sheet."""
    ws.merge_cells("A1:N1")
    ws["A1"] = "ASSESSMENT RESULTS\nDetailed assessment performance data"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Assessment_ID", 15),
        ("Employee_ID", 12),
        ("Module_ID", 15),
        ("Assessment_Type", 18),
        ("Date_Taken", 12),
        ("Score", 10),
        ("Pass_Threshold", 15),
        ("Pass_Fail", 10),
        ("Attempt_Number", 15),
        ("Time_Taken", 12),
        ("Questions_Correct", 18),
        ("Questions_Total", 15),
        ("Feedback_Provided", 18),
        ("Remediation_Required", 18),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Dropdowns
    dv_type = DataValidation(type="list", formula1='"Quiz,Practical,Simulation,Scenario,Tabletop"', allow_blank=False)
    dv_pass = DataValidation(type="list", formula1='"Pass,Fail"', allow_blank=False)
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)

    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_pass)
    ws.add_data_validation(dv_yesno)

    for r in range(4, 504):
        ws.cell(row=r, column=1, value=f"ASS-{r-3:04d}").font = Font(color="808080")

        for c in range(2, 15):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=4))
        dv_pass.add(ws.cell(row=r, column=8))
        dv_yesno.add(ws.cell(row=r, column=13))
        dv_yesno.add(ws.cell(row=r, column=14))

    ws.freeze_panes = "A4"


def create_simulation_results_sheet(ws, styles):
    """Create Simulation_Results sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = "SIMULATION RESULTS\nPhishing and behavioral simulation tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Campaign_ID", 15),
        ("Campaign_Name", 30),
        ("Campaign_Date", 12),
        ("Employee_ID", 12),
        ("Email_Sent", 10),
        ("Email_Opened", 12),
        ("Link_Clicked", 12),
        ("Credentials_Submitted", 18),
        ("Reported_Suspicious", 18),
        ("Time_To_Click", 15),
        ("Time_To_Report", 15),
        ("Remediation_Assigned", 18),
        ("Remediation_Completed", 18),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Dropdowns
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv_yesno)

    for r in range(4, 504):
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Add Yes/No dropdowns
        for col in [5, 6, 7, 8, 9, 12, 13]:
            dv_yesno.add(ws.cell(row=r, column=col))

    ws.freeze_panes = "A4"


def create_remediation_tracking_sheet(ws, styles):
    """Create Remediation_Tracking sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = "REMEDIATION TRACKING\nTrack remediation for failed assessments/simulations"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Remediation_ID", 15),
        ("Employee_ID", 12),
        ("Trigger_Event", 20),
        ("Trigger_Date", 12),
        ("Remediation_Level", 18),
        ("Remediation_Training", 35),
        ("Assigned_Date", 12),
        ("Due_Date", 12),
        ("Completion_Date", 15),
        ("Outcome", 15),
        ("Manager_Notified", 15),
        ("HR_Notified", 12),
        ("Notes", 50),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Dropdowns
    dv_trigger = DataValidation(type="list", formula1='"Failed Assessment,Clicked Phishing,Submitted Credentials,Pattern Failure"', allow_blank=False)
    dv_level = DataValidation(type="list", formula1='"Level 1,Level 2,Level 3"', allow_blank=False)
    dv_outcome = DataValidation(type="list", formula1='"Passed,Failed,Escalated,In Progress"', allow_blank=False)
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)

    ws.add_data_validation(dv_trigger)
    ws.add_data_validation(dv_level)
    ws.add_data_validation(dv_outcome)
    ws.add_data_validation(dv_yesno)

    for r in range(4, 254):
        ws.cell(row=r, column=1, value=f"REM-{r-3:03d}").font = Font(color="808080")

        for c in range(2, 14):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_trigger.add(ws.cell(row=r, column=3))
        dv_level.add(ws.cell(row=r, column=5))
        dv_outcome.add(ws.cell(row=r, column=10))
        dv_yesno.add(ws.cell(row=r, column=11))
        dv_yesno.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "A4"


def create_compliance_summary_sheet(ws, styles):
    """Create Compliance_Summary sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "COMPLIANCE SUMMARY\nAggregated compliance metrics by department/tier"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # By Department section
    ws["A3"] = "BY DEPARTMENT"
    ws["A3"].font = Font(bold=True, size=12)

    headers = [
        ("Department", 25),
        ("Total_Personnel", 15),
        ("Training_Required", 18),
        ("Completed", 12),
        ("Completion_Rate", 15),
        ("On_Time_Rate", 15),
        ("Average_Score", 15),
        ("Overdue_Count", 15),
        ("Compliance_Status", 18),
    ]

    row = 4
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for r in range(5, 25):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # By Training Tier section
    row = 27
    ws[f"A{row}"] = "BY TRAINING TIER"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    for col_idx, (header, _) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header.replace("Department", "Training_Tier"))
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    for r in range(row + 1, row + 10):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    ws.freeze_panes = "A5"


def create_overdue_alerts_sheet(ws, styles):
    """Create Overdue_Alerts sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "OVERDUE ALERTS\nReal-time overdue training alerts"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Employee_ID", 12),
        ("Employee_Name", 25),
        ("Manager", 25),
        ("Department", 20),
        ("Module_Title", 35),
        ("Due_Date", 12),
        ("Days_Overdue", 15),
        ("Escalation_Level", 18),
        ("Last_Reminder_Sent", 18),
        ("Action_Required", 40),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Escalation dropdown
    dv_escalation = DataValidation(type="list", formula1='"Level 1 (1-7 days),Level 2 (8-14 days),Level 3 (15-30 days),Level 4 (>30 days)"', allow_blank=False)
    ws.add_data_validation(dv_escalation)

    for r in range(4, 204):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_escalation.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"


def create_evidence_register_sheet(ws, styles):
    """Create Evidence_Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 40),
        ("Location", 40),
        ("Date_Collected", 15),
        ("Collected_By", 20),
        ("Status", 18),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_status = DataValidation(type="list", formula1='"Verified,Pending,Requires update,Not available"', allow_blank=False)
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_status.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "A4"


def create_dashboard_sheet(ws, styles):
    """Create Dashboard sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "TRAINING DELIVERY TRACKING - DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Completion Metrics"
    ws["A3"].font = Font(bold=True, size=12)

    metrics = [
        ("Total Training Records", '=COUNTA(Completion_Tracking!B4:B1003)'),
        ("Completed", '=COUNTIF(Completion_Tracking!I4:I1003,"Completed")'),
        ("In Progress", '=COUNTIF(Completion_Tracking!I4:I1003,"In Progress")'),
        ("Overdue", '=COUNTIF(Completion_Tracking!I4:I1003,"Overdue")'),
        ("Overall Completion Rate", '=IF(B4=0,"0%",ROUND(B5/B4*100,1)&"%")'),
    ]

    row = 4
    for label, formula in metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF", size=14)
        row += 1

    row += 2
    ws[f"A{row}"] = "Assessment Metrics"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    assessment_metrics = [
        ("Total Assessments", '=COUNTA(Assessment_Results!B4:B503)'),
        ("Pass Rate", '=IF(B11=0,"0%",ROUND(COUNTIF(Assessment_Results!H4:H503,"Pass")/B11*100,1)&"%")'),
        ("Average Score", '=AVERAGE(Assessment_Results!F4:F503)'),
    ]

    row += 1
    for label, formula in assessment_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    row += 2
    ws[f"A{row}"] = "Simulation Metrics"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    sim_metrics = [
        ("Total Simulations", '=COUNTA(Simulation_Results!D4:D503)'),
        ("Click Rate", '=IF(B17=0,"0%",ROUND(COUNTIF(Simulation_Results!G4:G503,"Yes")/B17*100,1)&"%")'),
        ("Report Rate", '=IF(B17=0,"0%",ROUND(COUNTIF(Simulation_Results!I4:I503,"Yes")/B17*100,1)&"%")'),
    ]

    row += 1
    for label, formula in sim_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    row += 2
    ws[f"A{row}"] = "Remediation Metrics"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    rem_metrics = [
        ("Active Remediation Cases", '=COUNTIF(Remediation_Tracking!J4:J253,"In Progress")'),
        ("Completed Remediation", '=COUNTIF(Remediation_Tracking!J4:J253,"Passed")'),
    ]

    row += 1
    for label, formula in rem_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def create_approval_signoff_sheet(ws, styles):
    """Create Approval_Sign_Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "MONTHLY ATTESTATION AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Reporting Period"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Report Period", ""),
        ("Total Personnel", "=Dashboard!B4"),
        ("Completion Rate", "=Dashboard!B8"),
        ("Overdue Count", "=Dashboard!B7"),
        ("Assessment Pass Rate", "=Dashboard!B12"),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    # Attestation sections
    sections = [
        ("PREPARED BY (Training Administrator)", "4472C4"),
        ("REVIEWED BY (HR Manager)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    for section_title, color in sections:
        row += 2
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = section_title
        ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

        fields = ["Name", "Date", "Signature"]
        row += 1
        for field in fields:
            ws[f"A{row}"] = field + ":"
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
            row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/11] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/11] Creating Personnel_Register sheet...")
        create_personnel_register_sheet(wb["Personnel_Register"], styles)

        logger.info("[3/11] Creating Completion_Tracking sheet...")
        create_completion_tracking_sheet(wb["Completion_Tracking"], styles)

        logger.info("[4/11] Creating Assessment_Results sheet...")
        create_assessment_results_sheet(wb["Assessment_Results"], styles)

        logger.info("[5/11] Creating Simulation_Results sheet...")
        create_simulation_results_sheet(wb["Simulation_Results"], styles)

        logger.info("[6/11] Creating Remediation_Tracking sheet...")
        create_remediation_tracking_sheet(wb["Remediation_Tracking"], styles)

        logger.info("[7/11] Creating Compliance_Summary sheet...")
        create_compliance_summary_sheet(wb["Compliance_Summary"], styles)

        logger.info("[8/11] Creating Overdue_Alerts sheet...")
        create_overdue_alerts_sheet(wb["Overdue_Alerts"], styles)

        logger.info("[9/11] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[10/11] Creating Dashboard sheet...")
        create_dashboard_sheet(wb["Dashboard"], styles)

        logger.info("[11/11] Creating Approval_Sign_Off sheet...")
        create_approval_signoff_sheet(wb["Approval_Sign_Off"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation following ISMS-IMP-A.6.3.3 specification
# =============================================================================
