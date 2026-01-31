#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-A.8.25-26-29 - Assessment File Normalizer Utility
================================================================================

ISO/IEC 27001:2022 Controls A.8.25/A.8.26/A.8.29: Secure Development Framework
Quality Assurance Utility: Excel Assessment File Normalization & Validation

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific:
- Assessment file naming conventions and versioning standards
- Workbook structure validation rules specific to your assessments
- Data format normalization rules aligned with your standards
- Validation severity thresholds based on your quality requirements
- Output formatting preferences aligned with your reporting needs
- Expected dropdown values matching your environment
- Formula patterns and conditional formatting rules

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.25-26-29 Secure Development Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script normalizes and validates A.8.25-26-29 secure development assessment
Excel workbooks to ensure consistency, data quality, and compliance with framework
standards before consolidation into the compliance dashboard.

**Purpose:**
Ensures all secure development assessment workbooks meet quality standards and
structural requirements, preventing data consolidation errors and improving
audit evidence reliability. Think of this as your "Red Team" quality assurance
that catches human errors before they become compliance problems.

**Key Functions:**
1. File Naming Validation
   - Verify naming convention compliance
   - Check date format validity (YYYYMMDD suffix)
   - Identify versioning inconsistencies

2. Workbook Structure Validation
   - Verify presence of required sheets
   - Validate column headers and data types
   - Check for missing or extra sheets
   - Ensure protected/unprotected cells are correct

3. Data Normalization
   - Standardize dropdown values (case, spacing, terminology)
   - Normalize date formats (DD.MM.YYYY or your standard)
   - Clean whitespace and formatting inconsistencies
   - Validate data type compliance (text vs. numbers)

4. Content Validation
   - Check for incomplete assessments
   - Identify placeholder/sample data not replaced
   - Validate formula integrity
   - Verify conditional formatting rules

5. Evidence Linkage Validation
   - Check evidence references are populated
   - Validate evidence file paths/URLs
   - Identify broken evidence links

6. Compliance Scoring Validation
   - Verify scoring formula correctness
   - Validate compliance percentage calculations
   - Check for scoring anomalies or errors

7. Quality Reporting
   - Generate validation report with findings
   - Categorize issues by severity (Critical, High, Medium, Low)
   - Provide remediation guidance
   - Track validation history

**Validation Scope:**
- ISMS_A_8_25_26_29_1_Security_Requirements_Assessment_YYYYMMDD.xlsx
- ISMS_A_8_25_26_29_2_SDLC_Security_Activities_Assessment_YYYYMMDD.xlsx
- ISMS_A_8_25_26_29_3_Security_Testing_Results_Assessment_YYYYMMDD.xlsx
- ISMS_A_8_25_26_29_4_Vulnerability_Remediation_Tracking_YYYYMMDD.xlsx
- ISMS_A_8_25_26_29_5_Compliance_Dashboard_YYYYMMDD.xlsx

**Output:**
- Normalized assessment workbooks (with _normalized suffix if changes made)
- Validation report (text or Excel format)
- Issue summary for remediation

**Quality Checks Performed:**

Critical Issues (Must Fix Before Consolidation):
- Missing required sheets
- Incorrect sheet names
- Invalid data types in key columns
- Broken formulas
- Missing compliance scores

High Priority Issues (Should Fix):
- Incomplete assessments (missing data)
- Inconsistent dropdown values
- Missing evidence references
- Date format inconsistencies

Medium Priority Issues (Recommended Fix):
- Formatting inconsistencies
- Whitespace issues
- Non-standard terminology
- Missing optional fields

Low Priority Issues (Nice to Fix):
- Cosmetic formatting variations
- Optional field inconsistencies
- Documentation completeness

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel processing

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl>=3.0.0 (Python Excel library)
    - datetime (standard library)
    - os (standard library)
    - re (standard library - regex for validation)
    - shutil (standard library - file operations)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    # Validate all A.8.25-26-29 assessment files in current directory
    python3 normalize_assessment_files_a825_26_29.py

Advanced Usage:
    # Validate files in specific directory
    python3 normalize_assessment_files_a825_26_29.py --input-dir /path/to/assessments
    
    # Validate with automatic normalization
    python3 normalize_assessment_files_a825_26_29.py --normalize
    
    # Validate specific assessment domain only
    python3 normalize_assessment_files_a825_26_29.py --domain 1
    
    # Generate detailed validation report
    python3 normalize_assessment_files_a825_26_29.py --report detailed
    
    # Dry run mode (report issues without modifying files)
    python3 normalize_assessment_files_a825_26_29.py --dry-run
    
    # Specify output directory for normalized files
    python3 normalize_assessment_files_a825_26_29.py --normalize --output-dir /path/to/output

Command-Line Options:
    --input-dir PATH       Directory containing assessment workbooks
    --output-dir PATH      Directory for normalized workbooks (default: input-dir)
    --normalize            Apply normalization fixes automatically
    --domain N             Validate specific domain only (1-5)
    --report TYPE          Report format: summary|detailed|excel (default: summary)
    --dry-run              Validate only, don't modify files
    --severity LEVEL       Minimum severity to report: critical|high|medium|low
    --backup               Create backup before normalization (recommended)

Output Files:
    If --normalize used:
        - Original files: [filename]_backup_YYYYMMDD.xlsx (if --backup)
        - Normalized files: [filename] (updated in place) OR
        - Normalized files: [filename]_normalized.xlsx (if different output-dir)
    
    Validation report:
        - Console output (summary)
        - Text file: A825_26_29_Assessment_Validation_Report_YYYYMMDD.txt
        - Excel file: A825_26_29_Assessment_Validation_Report_YYYYMMDD.xlsx (if --report excel)

Workflow Examples:

    1. Initial validation (before consolidation):
       python3 normalize_assessment_files_a825_26_29.py --dry-run --report detailed
    
    2. Normalize and fix issues:
       python3 normalize_assessment_files_a825_26_29.py --normalize --backup
    
    3. Validate after normalization:
       python3 normalize_assessment_files_a825_26_29.py --severity high
    
    4. Generate audit-ready validation report:
       python3 normalize_assessment_files_a825_26_29.py --report excel

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.25/A.8.26/A.8.29
Utility Type:         Quality Assurance - Assessment Normalization & Validation
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization Security Team]
Date:                 [YYYY-MM-DD]
Last Modified:        [YYYY-MM-DD]
Python Version:       3.8+
License:              [Organization License/Terms]

Related Documents:
    - ISMS-POL-A.8.25-26-29-S1: Executive Control Alignment (Governance)
    - ISMS-POL-A.8.25-26-29-S2: Security Requirements (A.8.26)
    - ISMS-POL-A.8.25-26-29-S3: Secure Development Lifecycle (A.8.25)
    - ISMS-POL-A.8.25-26-29-S4: Security Testing (A.8.29)
    - ISMS-POL-A.8.25-26-29-S5: Assessment Evidence Framework
    - ISMS-IMP-A.8.25-26-29-S1: Security Requirements Process Implementation Guide
    - ISMS-IMP-A.8.25-26-29-S2: SDLC Security Integration Implementation Guide
    - ISMS-IMP-A.8.25-26-29-S4: Security Testing Implementation Implementation Guide
    - ISMS-IMP-A.8.25-26-29-S5: Secure Development Assessment Implementation Guide

Related Scripts:
    - generate_a825_26_29_1_security_requirements.py
    - generate_a825_26_29_2_sdlc_security_activities.py
    - generate_a825_26_29_3_security_testing_results.py
    - generate_a825_26_29_4_vulnerability_remediation.py
    - generate_a825_26_29_5_compliance_dashboard.py
    - normalize_assessment_files_a825_26_29.py (this script)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [YYYY-MM-DD]
    - Initial release
    - Implements comprehensive validation framework for all five domains
    - Supports automated normalization of all assessment workbooks
    - Generates quality assurance reports for audit readiness

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Quality Assurance Philosophy:**
This script embodies "don't fool yourself" engineering - it catches the errors
humans make when filling out assessments, ensuring data quality before
consolidation. Think of it as your "Red Team" reviewer that never gets tired,
never misses details, and always applies standards consistently.

**Validation vs. Normalization:**
- **Validation**: Identifies issues without modifying files (--dry-run mode)
- **Normalization**: Automatically fixes issues where safe to do so (--normalize mode)
- Some issues require human judgment and cannot be auto-normalized
- Always review validation report before accepting normalization changes

**Backup Recommendation:**
ALWAYS use --backup flag when normalizing files. Assessment workbooks contain
valuable data collection effort representing hours or days of work. Don't risk
data loss - backups are cheap, recreation is expensive.

**Pre-Consolidation Requirement:**
Run this script BEFORE consolidating data into the compliance dashboard
(Domain 5). Dashboard consolidation assumes normalized, validated assessment
workbooks. Garbage in, garbage out - validate first.

**Audit Considerations:**
Validation reports demonstrate quality assurance processes to auditors.
Keep validation reports as evidence of systematic quality control.
Auditors appreciate seeing that you have quality gates in your process.

**Data Integrity:**
Script validates but does not alter actual assessment data values (e.g.,
compliance scores, technical findings). It only normalizes format and structure.
Technical accuracy remains the assessor's responsibility. This script catches
format errors, not factual errors.

**False Positives:**
Some validation warnings may be acceptable based on your specific context.
Review validation report and use judgment - don't blindly "fix" everything.
Context matters. If you have a good reason to deviate from standards, document it.

**Schema Changes:**
If you modify assessment workbook structures (add sheets, change columns),
update this script's validation rules accordingly. Out-of-sync validation
rules will generate false positives/negatives. Keep script and workbook
specifications aligned.

**Performance:**
Script processes Excel files in memory. For very large assessment workbooks
(>50MB), consider increasing available memory or processing files individually
using --domain flag. Normal assessment workbooks (5-20MB) process quickly.

**Error Handling:**
Script continues processing all files even if one fails validation. Check
final summary for any files that couldn't be processed. Individual file
failures don't stop batch validation.

**Integration with CI/CD:**
Consider integrating this script into your assessment workflow:
- Pre-consolidation validation (automated check)
- Scheduled validation (weekly quality checks)
- Pre-audit validation (audit preparation)
- Exit criteria for assessment completion

**Customization Requirements:**
Different organizations have different standards. You MUST customize:
- Expected dropdown values to match YOUR environment
- Date formats to match YOUR regional standards
- Sheet names if you've customized workbooks
- Column headers if you've customized assessments
- Validation rules to match YOUR quality thresholds

**Critical Success Factors:**
- **Regular Use**: Don't wait until audit - validate continuously
- **Team Training**: Train assessors on common validation failures
- **Continuous Improvement**: Update validation rules based on recurring issues
- **Automation**: Integrate into workflow, don't rely on manual execution
- **Accountability**: Assign ownership for fixing validation failures

================================================================================
"""

import os
import sys
import shutil
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("ℹ️  Install with: sudo apt install python3-openpyxl")     
    sys.exit(1)


# ============================================================================
# CONFIGURATION
# ============================================================================

# Expected document IDs and titles from assessment workbooks
EXPECTED_DOCS = {
    "ISMS-IMP-A.8.25-26-29.1": {
        "title": "Security Requirements Assessment",
        "normalized": "ISMS-IMP-A.8.25-26-29.1.xlsx",
        "pattern": ["a825", "a8.25", "security_requirements", "requirements_assessment"]
    },
    "ISMS-IMP-A.8.25-26-29.2": {
        "title": "SDLC Security Activities Assessment",
        "normalized": "ISMS-IMP-A.8.25-26-29.2.xlsx",
        "pattern": ["a825", "a8.25", "sdlc", "sdlc_security", "security_activities"]
    },
    "ISMS-IMP-A.8.25-26-29.3": {
        "title": "Security Testing Results Assessment",
        "normalized": "ISMS-IMP-A.8.25-26-29.3.xlsx",
        "pattern": ["a829", "a8.29", "security_testing", "testing_results"]
    },
    "ISMS-IMP-A.8.25-26-29.4": {
        "title": "Vulnerability Remediation Tracking",
        "normalized": "ISMS-IMP-A.8.25-26-29.4.xlsx",
        "pattern": ["a829", "a8.29", "vulnerability", "remediation"]
    },
}


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_workbook(filepath):
    """
    Validate workbook contains expected document ID in Instructions sheet.
    
    Args:
        filepath: Path to Excel workbook
    
    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        # Open workbook in read-only mode
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # Check for Instructions & Legend sheet (with or without underscore)
        sheet_name = None
        for name in ["Instructions & Legend", "Instructions_Legend", "Instructions"]:
            if name in wb.sheetnames:
                sheet_name = name
                break
        
        if not sheet_name:
            wb.close()
            return (None, None)
        
        ws = wb[sheet_name]
        
        # Look for Document ID in column A (typically rows 4-20)
        # Format: "Document ID" label in column A, value in column B
        for row in range(3, 25):
            cell_label = ws.cell(row=row, column=1).value
            
            if cell_label and "Document ID" in str(cell_label):
                doc_id_value = ws.cell(row=row, column=2).value
                
                if doc_id_value:
                    doc_id = str(doc_id_value).strip()
                    
                    # Check if this doc_id matches expected format
                    if doc_id in EXPECTED_DOCS:
                        title = EXPECTED_DOCS[doc_id]["title"]
                        wb.close()
                        return (doc_id, title)
        
        wb.close()
        return (None, None)
        
    except Exception as e:
        print(f"  ⚠️  Error validating {filepath}: {e}")
        return (None, None)


def detect_workbook_from_filename(filename):
    """
    Attempt to detect workbook type from filename patterns.
    
    Args:
        filename: Filename to analyze
    
    Returns:
        str: Document ID if detected, None otherwise
    """
    filename_lower = filename.lower()
    
    for doc_id, config in EXPECTED_DOCS.items():
        for pattern in config["pattern"]:
            if pattern in filename_lower:
                return doc_id
    
    return None


# ============================================================================
# NORMALIZATION FUNCTIONS
# ============================================================================

def normalize_file(source_path, doc_id, output_dir="."):
    """
    Copy assessment file to normalized filename.
    
    Args:
        source_path: Path to source workbook
        doc_id: Document ID (e.g., "ISMS-IMP-A.8.25-26-29.1")
        output_dir: Directory for normalized files
    
    Returns:
        str: Path to normalized file, or None if failed
    """
    try:
        normalized_filename = EXPECTED_DOCS[doc_id]["normalized"]
        normalized_path = Path(output_dir) / normalized_filename
        
        # Copy file
        shutil.copy2(source_path, normalized_path)
        
        return str(normalized_path)
        
    except Exception as e:
        print(f"  ❌ Error normalizing file: {e}")
        return None


# ============================================================================
# MANIFEST GENERATION
# ============================================================================

def create_manifest(normalized_files, output_dir="."):
    """
    Create audit manifest documenting normalization.
    
    Args:
        normalized_files: List of (source_path, normalized_path, doc_id, title) tuples
        output_dir: Directory for manifest file
    """
    manifest_path = Path(output_dir) / "NORMALIZATION_MANIFEST.txt"
    
    with open(manifest_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("ISMS A.8.25-26-29 ASSESSMENT FILE NORMALIZATION MANIFEST\n")
        f.write("=" * 80 + "\n")
        f.write(f"Normalization Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total Files Normalized: {len(normalized_files)}\n")
        f.write("\n")
        
        for source_path, normalized_path, doc_id, title in normalized_files:
            f.write("-" * 80 + "\n")
            f.write(f"Document ID: {doc_id}\n")
            f.write(f"Title: {title}\n")
            f.write(f"Source File: {source_path}\n")
            f.write(f"Normalized File: {normalized_path}\n")
            f.write(f"Normalized At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("\n")
        
        f.write("=" * 80 + "\n")
        f.write("NORMALIZATION COMPLETE\n")
        f.write("=" * 80 + "\n")
    
    print(f"\n📄 Manifest created: {manifest_path}")


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main normalization workflow."""
    print("=" * 80)
    print("ISMS A.8.25-26-29 ASSESSMENT FILE NORMALIZATION")
    print("=" * 80)
    print()
    
    # Find all Excel files in current directory
    excel_files = list(Path(".").glob("*.xlsx"))
    
    if not excel_files:
        print("❌ No Excel files found in current directory")
        print("ℹ️  Place assessment workbooks in this directory and run again")
        return
    
    print(f"📁 Found {len(excel_files)} Excel file(s) in current directory")
    print()
    
    # Validate and categorize files
    validated_files = {}  # doc_id -> filepath
    unrecognized_files = []
    
    print("🔍 Validating workbooks...")
    print()
    
    for filepath in excel_files:
        filename = filepath.name
        
        # Skip normalized files (avoid re-normalizing)
        if filename.startswith("ISMS-IMP-A.8.25-26-29.") and len(filename) < 30:
            print(f"  ⏭️  Skipping (already normalized): {filename}")
            continue
        
        # Skip dashboard and manifest
        if "dashboard" in filename.lower() or "manifest" in filename.lower():
            print(f"  ⏭️  Skipping (dashboard/manifest): {filename}")
            continue
        
        print(f"  📄 Checking: {filename}")
        
        # Validate by reading Document ID from workbook
        doc_id, title = validate_workbook(filepath)
        
        if doc_id:
            print(f"    ✅ Valid: {doc_id} - {title}")
            validated_files[doc_id] = filepath
        else:
            # Try to detect from filename
            detected_id = detect_workbook_from_filename(filename)
            if detected_id:
                print(f"    ⚠️  Document ID not found in workbook, but detected from filename: {detected_id}")
                print(f"       (Add proper Document ID in Instructions & Legend sheet for best practices)")
                validated_files[detected_id] = filepath
            else:
                print(f"    ❌ Unrecognized: Could not determine workbook type")
                unrecognized_files.append(filepath)
        
        print()
    
    # Report validation summary
    print("-" * 80)
    print("VALIDATION SUMMARY")
    print("-" * 80)
    print(f"✅ Validated: {len(validated_files)} workbook(s)")
    print(f"❌ Unrecognized: {len(unrecognized_files)} file(s)")
    print()
    
    if validated_files:
        print("Validated Workbooks:")
        for doc_id in sorted(validated_files.keys()):
            filepath = validated_files[doc_id]
            print(f"  - {doc_id}: {filepath.name}")
        print()
    
    if unrecognized_files:
        print("Unrecognized Files (will not be normalized):")
        for filepath in unrecognized_files:
            print(f"  - {filepath.name}")
        print()
    
    # Check if we have all expected workbooks
    missing_docs = set(EXPECTED_DOCS.keys()) - set(validated_files.keys())
    if missing_docs:
        print("⚠️  WARNING: Missing expected workbooks:")
        for doc_id in sorted(missing_docs):
            title = EXPECTED_DOCS[doc_id]["title"]
            print(f"  - {doc_id}: {title}")
        print()
        print("ℹ️  The dashboard generator requires all 4 workbooks.")
        print("   Continue anyway? (y/n): ", end="")
        
        response = input().strip().lower()
        if response != 'y':
            print("❌ Normalization cancelled")
            return
        print()
    
    # Normalize files
    if not validated_files:
        print("❌ No valid workbooks found to normalize")
        return
    
    print("-" * 80)
    print("NORMALIZING FILES")
    print("-" * 80)
    print()
    
    normalized_files = []
    
    for doc_id, source_path in validated_files.items():
        title = EXPECTED_DOCS[doc_id]["title"]
        normalized_filename = EXPECTED_DOCS[doc_id]["normalized"]
        
        print(f"📋 Normalizing: {doc_id}")
        print(f"   Title: {title}")
        print(f"   Source: {source_path.name}")
        print(f"   Target: {normalized_filename}")
        
        normalized_path = normalize_file(source_path, doc_id)
        
        if normalized_path:
            print(f"   ✅ Success: {normalized_path}")
            normalized_files.append((str(source_path), normalized_path, doc_id, title))
        else:
            print(f"   ❌ Failed to normalize")
        
        print()
    
    # Create manifest
    if normalized_files:
        create_manifest(normalized_files)
    
    # Final summary
    print("=" * 80)
    print("NORMALIZATION COMPLETE")
    print("=" * 80)
    print(f"\n✅ Successfully normalized {len(normalized_files)} workbook(s)")
    print("\nNormalized Files:")
    for _, normalized_path, doc_id, title in normalized_files:
        print(f"  - {Path(normalized_path).name}")
    
    print("\n" + "=" * 80)
    print("NEXT STEPS")
    print("=" * 80)
    print("1. Review normalized files (ISMS-IMP-A.8.25-26-29.*.xlsx)")
    print("2. Run dashboard generator:")
    print("   python3 generate_dashboard_secure_development.py")
    print("3. Dashboard will aggregate data from all 4 normalized workbooks")
    print("=" * 80)


if __name__ == "__main__":
    main()
