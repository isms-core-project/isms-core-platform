#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
A.5.17 Sanity Check - Validates all A.5.17 Authentication Information workbooks.

Usage: python3 sanity_check_a517.py
"""

import logging
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

CONTROL_ID = "A.5.17"
WORKBOOK_DIR = Path(__file__).parent.parent / "90_workbooks"

EXPECTED_WORKBOOKS = {
    "A.5.17.1": {
        "pattern": "ISMS-IMP-A.5.17.1_Authentication_Policy*.xlsx",
        "required_sheets": ["Instructions", "Password_Policy", "MFA_Requirements", "Credential_Standards", "User_Responsibilities", "System_Requirements", "Evidence_Register", "Approval_SignOff"]
    },
    "A.5.17.2": {
        "pattern": "ISMS-IMP-A.5.17.2_Credential_Lifecycle*.xlsx",
        "required_sheets": ["Instructions", "Allocation_Process", "Change_Management", "Recovery_Process", "Revocation_Process", "Audit_Log_Requirements", "Evidence_Register", "Approval_SignOff"]
    },
    "A.5.17.3": {
        "pattern": "ISMS-IMP-A.5.17.3_Password_System*.xlsx",
        "required_sheets": ["Instructions", "System_Inventory", "Security_Assessment", "Storage_Assessment", "Integration_Assessment", "Gap_Analysis", "Evidence_Register", "Approval_SignOff"]
    },
    "A.5.17.4": {
        "pattern": "ISMS-IMP-A.5.17.4_Compliance*.xlsx",
        "required_sheets": ["Instructions", "Executive_Summary", "Compliance_KPIs", "Authentication_Events", "Audit_Findings", "Remediation_Tracker", "Evidence_Register", "Approval_SignOff"]
    },
}


def check_workbook(filepath: Path, expected_sheets: list) -> tuple:
    issues = []
    passed = failed = 0
    try:
        wb = load_workbook(filepath, data_only=False)
        for sheet in expected_sheets:
            if sheet in wb.sheetnames:
                passed += 1
            else:
                failed += 1
                issues.append(f"Missing sheet: {sheet}")
        for ws in wb.worksheets:
            if ws["A1"].value is None:
                issues.append(f"Sheet '{ws.title}' has no header")
                failed += 1
            else:
                passed += 1
            if ws.data_validations.dataValidation:
                passed += 1
            elif ws.title != "Instructions":
                issues.append(f"Sheet '{ws.title}' has no data validations")
        wb.close()
    except Exception as e:
        failed += 1
        issues.append(f"Error: {e}")
    return passed, failed, issues


def main() -> int:
    logger.info("=" * 70)
    logger.info(f"{CONTROL_ID} Sanity Check")
    logger.info("=" * 70)

    if not WORKBOOK_DIR.exists():
        logger.warning(f"Workbook directory not found")
        return 0

    total_passed = total_failed = 0
    all_issues = []

    for doc_id, config in EXPECTED_WORKBOOKS.items():
        logger.info(f"\nChecking {doc_id}...")
        workbooks = list(WORKBOOK_DIR.glob(config["pattern"]))
        if not workbooks:
            logger.warning(f"  No workbooks found matching {config['pattern']}")
            total_failed += 1
            all_issues.append(f"{doc_id}: No workbook found")
            continue

        for wb_path in workbooks:
            logger.info(f"  Checking: {wb_path.name}")
            passed, failed, issues = check_workbook(wb_path, config["required_sheets"])
            total_passed += passed
            total_failed += failed
            for issue in issues:
                logger.warning(f"    ⚠ {issue}")
                all_issues.append(f"{doc_id}: {issue}")
            if not issues:
                logger.info(f"    ✓ All checks passed ({passed} checks)")

    logger.info("\n" + "=" * 70)
    logger.info("SANITY CHECK SUMMARY")
    logger.info("=" * 70)
    logger.info(f"Total checks passed: {total_passed}")
    logger.info(f"Total checks failed: {total_failed}")
    if all_issues:
        logger.info(f"\nIssues found ({len(all_issues)}):")
        for issue in all_issues:
            logger.info(f"  - {issue}")
    else:
        logger.info("\n✓ All sanity checks passed!")
    logger.info("=" * 70)
    return 0 if total_failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.17 Authentication Information control
# =============================================================================
