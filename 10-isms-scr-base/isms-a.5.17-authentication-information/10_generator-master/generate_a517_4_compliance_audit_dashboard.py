#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
A.5.17.4 Compliance and Audit Dashboard Generator
================================================================================

Generates Excel workbook for monitoring authentication compliance, tracking
KPIs, audit findings, and remediation progress per ISO 27001:2022 A.5.17.

Sheets:
    1. Instructions - Completion guidance
    2. Executive_Summary - High-level compliance metrics
    3. Compliance_KPIs - Key performance indicators
    4. Authentication_Events - Event monitoring summary
    5. Audit_Findings - Audit finding tracking
    6. Remediation_Tracker - Gap remediation progress
    7. Evidence_Register - Supporting documentation
    8. Approval_SignOff - Dashboard approval

Usage:
    python3 generate_a517_4_compliance_audit_dashboard.py
================================================================================
"""

import logging
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

DOCUMENT_ID = "ISMS-IMP-A.5.17.4"
WORKBOOK_NAME = "Compliance and Audit Dashboard"
CONTROL_ID = "A.5.17"
CONTROL_NAME = "Authentication Information"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
METRIC_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)
METRIC_FONT = Font(name="Calibri", size=16, bold=True)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP_LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


def set_column_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_header_row(ws, row: int, headers: list):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def add_data_validation(ws, cell_range: str, formula: str):
    dv = DataValidation(type="list", formula1=formula, showDropDown=False, allowBlank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def create_metric_box(ws, row: int, col: int, label: str, value: str = "", fill=METRIC_FILL):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    cell = ws.cell(row=row, column=col, value=label)
    cell.font = BOLD_FONT
    cell.fill = fill
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    value_cell = ws.cell(row=row+1, column=col, value=value)
    value_cell.font = METRIC_FONT
    value_cell.fill = INPUT_FILL
    value_cell.alignment = CENTER_ALIGN
    value_cell.border = THIN_BORDER
    ws.row_dimensions[row+1].height = 35


def create_instructions_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Instructions"

    ws.merge_cells("A1:H1")
    ws["A1"].value = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    metadata = [
        ("Document ID:", DOCUMENT_ID),
        ("Control Reference:", CONTROL_REF),
        ("Generated Date:", GENERATED_DATE),
        ("Version:", "1.0"),
    ]

    row = 3
    for label, value in metadata:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="PURPOSE").font = BOLD_FONT
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1).value = (
        "This dashboard provides visibility into authentication compliance per ISO 27001:2022 A.5.17. "
        "It tracks KPIs, authentication events, audit findings, and remediation progress. Use for "
        "management reporting, audit preparation, and continuous improvement of authentication controls."
    )
    ws.cell(row=row, column=1).font = NORMAL_FONT
    ws.cell(row=row, column=1).alignment = TOP_LEFT_ALIGN
    ws.row_dimensions[row].height = 50

    row += 2
    ws.cell(row=row, column=1, value="SHEET DESCRIPTIONS").font = BOLD_FONT
    row += 1

    sheets = [
        ("Executive_Summary", "High-level compliance overview with key metrics"),
        ("Compliance_KPIs", "Key performance indicators for authentication"),
        ("Authentication_Events", "Authentication event monitoring summary"),
        ("Audit_Findings", "Audit finding tracking"),
        ("Remediation_Tracker", "Gap remediation progress"),
        ("Evidence_Register", "Supporting documentation"),
        ("Approval_SignOff", "Dashboard approval"),
    ]

    for sheet_name, description in sheets:
        ws.cell(row=row, column=1, value=sheet_name).font = BOLD_FONT
        ws.cell(row=row, column=2, value=description).font = NORMAL_FONT
        row += 1

    set_column_widths(ws, {"A": 25, "B": 50, "C": 20, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20})
    ws.freeze_panes = "A3"
    logger.info("Created Instructions sheet")


def create_executive_summary_sheet(wb: Workbook):
    ws = wb.create_sheet("Executive_Summary")

    ws.merge_cells("A1:J1")
    ws["A1"].value = "Authentication Compliance - Executive Summary"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    ws.cell(row=3, column=1, value="Reporting Period:").font = BOLD_FONT
    ws.cell(row=3, column=2).fill = INPUT_FILL
    ws.cell(row=3, column=3, value="Report Date:").font = BOLD_FONT
    ws.cell(row=3, column=4, value=GENERATED_DATE).font = NORMAL_FONT

    row = 5
    create_metric_box(ws, row, 1, "Overall Compliance")
    create_metric_box(ws, row, 3, "MFA Adoption Rate")
    create_metric_box(ws, row, 5, "Password Policy Compliance")
    create_metric_box(ws, row, 7, "Failed Login Rate")
    create_metric_box(ws, row, 9, "Open Findings")

    row = 8
    create_metric_box(ws, row, 1, "Systems Assessed", fill=GREEN_FILL)
    create_metric_box(ws, row, 3, "Users with MFA", fill=GREEN_FILL)
    create_metric_box(ws, row, 5, "Compliant Passwords", fill=GREEN_FILL)
    create_metric_box(ws, row, 7, "Avg Daily Lockouts", fill=YELLOW_FILL)
    create_metric_box(ws, row, 9, "Critical Gaps", fill=RED_FILL)

    row = 12
    ws.cell(row=row, column=1, value="Compliance by Area").font = BOLD_FONT
    ws.merge_cells(f"A{row}:E{row}")

    row += 1
    headers = ["Control Area", "Total Requirements", "Compliant", "Non-Compliant", "% Compliant"]
    create_header_row(ws, row, headers)

    row += 1
    areas = [
        ("Password Policy", "", "", "", ""),
        ("MFA Implementation", "", "", "", ""),
        ("Credential Lifecycle", "", "", "", ""),
        ("System Security", "", "", "", ""),
        ("User Responsibilities", "", "", "", ""),
        ("Audit Logging", "", "", "", ""),
        ("TOTAL", "", "", "", ""),
    ]

    for area in areas:
        for col, value in enumerate(area, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = BOLD_FONT if area[0] == "TOTAL" else NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if col > 1 else LEFT_ALIGN
            if col > 1:
                cell.fill = INPUT_FILL
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Key Actions Required").font = BOLD_FONT
    ws.merge_cells(f"A{row}:J{row}")

    row += 1
    action_headers = ["Priority", "Action Item", "Owner", "Due Date", "Status", "", "", "", "", ""]
    create_header_row(ws, row, action_headers)

    row += 1
    for i in range(5):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"A{row-5}:A{row-1}", '"Critical,High,Medium,Low"')
    add_data_validation(ws, f"E{row-5}:E{row-1}", '"Not Started,In Progress,Completed,Blocked"')

    set_column_widths(ws, {"A": 20, "B": 30, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14})
    ws.freeze_panes = "A3"
    logger.info("Created Executive_Summary sheet")


def create_compliance_kpis_sheet(wb: Workbook):
    ws = wb.create_sheet("Compliance_KPIs")

    ws.merge_cells("A1:I1")
    ws["A1"].value = "Authentication Compliance KPIs"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["KPI ID", "KPI Name", "Description", "Target", "Current", "Status", "Trend", "Owner", "Notes"]
    create_header_row(ws, 3, headers)

    kpis = [
        ("KPI-AUTH-001", "MFA Enrollment Rate", "% of users enrolled in MFA", "100%", "", "", "", "", ""),
        ("KPI-AUTH-002", "MFA Usage Rate", "% of logins using MFA", "≥95%", "", "", "", "", ""),
        ("KPI-AUTH-003", "Password Policy Compliance", "% of accounts meeting password policy", "100%", "", "", "", "", ""),
        ("KPI-AUTH-004", "Failed Login Rate", "Failed logins as % of total attempts", "<5%", "", "", "", "", ""),
        ("KPI-AUTH-005", "Account Lockout Rate", "Daily account lockouts per 1000 users", "<2", "", "", "", "", ""),
        ("KPI-AUTH-006", "Password Reset Volume", "Self-service resets per 1000 users/month", "<50", "", "", "", "", ""),
        ("KPI-AUTH-007", "Credential Exposure", "Accounts found in breach databases", "0", "", "", "", "", ""),
        ("KPI-AUTH-008", "Service Account Rotation", "% of service accounts rotated on schedule", "100%", "", "", "", "", ""),
        ("KPI-AUTH-009", "SSO Coverage", "% of applications integrated with SSO", "≥90%", "", "", "", "", ""),
        ("KPI-AUTH-010", "Privileged Account MFA", "% of privileged accounts with MFA", "100%", "", "", "", "", ""),
        ("KPI-AUTH-011", "Training Completion", "% of users completing auth security training", "≥95%", "", "", "", "", ""),
        ("KPI-AUTH-012", "Audit Finding Closure", "% of findings closed within SLA", "≥85%", "", "", "", "", ""),
    ]

    row = 4
    for kpi in kpis:
        for col, value in enumerate(kpi, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [5, 6, 7, 8, 9]:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"F4:F{row+5}", '"On Target,At Risk,Below Target,Not Measured"')
    add_data_validation(ws, f"G4:G{row+5}", '"↑ Improving,→ Stable,↓ Declining,New"')

    set_column_widths(ws, {"A": 14, "B": 25, "C": 38, "D": 12, "E": 12, "F": 14, "G": 12, "H": 16, "I": 22})
    ws.freeze_panes = "A4"
    logger.info("Created Compliance_KPIs sheet")


def create_authentication_events_sheet(wb: Workbook):
    ws = wb.create_sheet("Authentication_Events")

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Authentication Event Monitoring"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["Event Category", "This Period", "Previous Period", "Change", "Threshold", "Alert Status", "Investigation", "Notes"]
    create_header_row(ws, 3, headers)

    events = [
        ("Total Login Attempts", "", "", "", "Baseline", "", "", ""),
        ("Successful Logins", "", "", "", "Baseline", "", "", ""),
        ("Failed Logins", "", "", "", "<5% of total", "", "", ""),
        ("Account Lockouts", "", "", "", "<2 per 1000 users/day", "", "", ""),
        ("Password Resets (Self-Service)", "", "", "", "Monitor trend", "", "", ""),
        ("Password Resets (Admin)", "", "", "", "Alert on spike", "", "", ""),
        ("MFA Failures", "", "", "", "<2% of MFA attempts", "", "", ""),
        ("MFA Bypass Events", "", "", "", "0 (alert on any)", "", "", ""),
        ("Impossible Travel Alerts", "", "", "", "Review all", "", "", ""),
        ("Unknown Device Logins", "", "", "", "Review all new", "", "", ""),
        ("After Hours Logins", "", "", "", "Review anomalies", "", "", ""),
        ("Service Account Logins", "", "", "", "Baseline", "", "", ""),
        ("Privilege Escalation", "", "", "", "Review all", "", "", ""),
        ("Emergency Access (Break-Glass)", "", "", "", "0 normal", "", "", ""),
    ]

    row = 4
    for event in events:
        for col, value in enumerate(event, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [2, 3, 4, 6, 7, 8]:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"F4:F{row+5}", '"Normal,Warning,Alert,Critical"')
    add_data_validation(ws, f"G4:G{row+5}", '"None Required,In Progress,Completed,Escalated"')

    set_column_widths(ws, {"A": 28, "B": 14, "C": 16, "D": 12, "E": 22, "F": 14, "G": 14, "H": 25})
    ws.freeze_panes = "A4"
    logger.info("Created Authentication_Events sheet")


def create_audit_findings_sheet(wb: Workbook):
    ws = wb.create_sheet("Audit_Findings")

    ws.merge_cells("A1:K1")
    ws["A1"].value = "Authentication Audit Findings"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["Finding ID", "Audit Type", "Audit Date", "Finding Description", "Affected Area", "Severity", "Recommendation", "Owner", "Due Date", "Status", "Notes"]
    create_header_row(ws, 3, headers)

    row = 4
    for i in range(15):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"B4:B{row-1}", '"Internal Audit,ISO 27001,Penetration Test,SOC 2,Customer Audit,Self-Assessment,Other"')
    add_data_validation(ws, f"F4:F{row-1}", '"Critical,High,Medium,Low,Observation"')
    add_data_validation(ws, f"J4:J{row-1}", '"Open,In Progress,Remediated,Verified,Risk Accepted,Overdue"')

    set_column_widths(ws, {"A": 12, "B": 14, "C": 12, "D": 32, "E": 18, "F": 10, "G": 28, "H": 16, "I": 12, "J": 12, "K": 22})
    ws.freeze_panes = "A4"
    logger.info("Created Audit_Findings sheet")


def create_remediation_tracker_sheet(wb: Workbook):
    ws = wb.create_sheet("Remediation_Tracker")

    ws.merge_cells("A1:J1")
    ws["A1"].value = "Remediation Tracker"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["Item ID", "Source", "Description", "Priority", "Remediation Plan", "Owner", "Start Date", "Target Date", "Status", "Notes"]
    create_header_row(ws, 3, headers)

    row = 4
    for i in range(15):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"B4:B{row-1}", '"Audit Finding,Assessment Gap,Incident RCA,Risk Assessment,Management Review,Other"')
    add_data_validation(ws, f"D4:D{row-1}", '"Critical,High,Medium,Low"')
    add_data_validation(ws, f"I4:I{row-1}", '"Not Started,Planning,In Progress,Testing,Completed,On Hold"')

    set_column_widths(ws, {"A": 12, "B": 16, "C": 32, "D": 10, "E": 32, "F": 16, "G": 12, "H": 12, "I": 12, "J": 25})
    ws.freeze_panes = "A4"
    logger.info("Created Remediation_Tracker sheet")


def create_evidence_register_sheet(wb: Workbook):
    ws = wb.create_sheet("Evidence_Register")

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Evidence Register - Compliance Dashboard"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["Evidence ID", "Evidence Type", "Description", "Related Section", "Location/Link", "Date", "Collected By", "Status"]
    create_header_row(ws, 3, headers)

    evidence = [
        ("EV-517-CD-001", "Report", "Monthly compliance dashboard", "Executive_Summary", "", "", "", ""),
        ("EV-517-CD-002", "Metrics", "KPI data extract", "Compliance_KPIs", "", "", "", ""),
        ("EV-517-CD-003", "Log Extract", "Authentication event summary", "Authentication_Events", "", "", "", ""),
        ("EV-517-CD-004", "Audit Report", "Latest audit report", "Audit_Findings", "", "", "", ""),
        ("EV-517-CD-005", "Status Report", "Remediation progress report", "Remediation_Tracker", "", "", "", ""),
    ]

    row = 4
    for item in evidence:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [5, 6, 7, 8]:
                cell.fill = INPUT_FILL
        row += 1

    for i in range(10):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"H4:H{row-1}", '"Pending,Collected,Verified,Expired,N/A"')

    set_column_widths(ws, {"A": 16, "B": 16, "C": 32, "D": 20, "E": 28, "F": 12, "G": 16, "H": 12})
    ws.freeze_panes = "A4"
    logger.info("Created Evidence_Register sheet")


def create_approval_signoff_sheet(wb: Workbook):
    ws = wb.create_sheet("Approval_SignOff")

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Dashboard Approval and Sign-Off"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    ws.cell(row=3, column=1, value="Document Information").font = BOLD_FONT

    info = [
        ("Document ID:", DOCUMENT_ID, "Version:", "1.0"),
        ("Document Title:", WORKBOOK_NAME, "Date:", GENERATED_DATE),
        ("Control Reference:", CONTROL_ID, "Reporting Period:", ""),
    ]

    row = 4
    for item in info:
        ws.cell(row=row, column=1, value=item[0]).font = BOLD_FONT
        ws.cell(row=row, column=2, value=item[1]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=item[2]).font = BOLD_FONT
        cell = ws.cell(row=row, column=4, value=item[3])
        cell.font = NORMAL_FONT
        if item[3] == "":
            cell.fill = INPUT_FILL
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Approval Signatures").font = BOLD_FONT

    row += 1
    headers = ["Role", "Name", "Signature", "Date", "Status", "Comments"]
    create_header_row(ws, row, headers)

    row += 1
    approvers = [
        ("Dashboard Owner", "", "", "", "", ""),
        ("IT Security Manager", "", "", "", "", ""),
        ("Information Security Officer", "", "", "", "", ""),
        ("CISO", "", "", "", "", ""),
    ]

    for approver in approvers:
        for col, value in enumerate(approver, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col > 1:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"E{row-4}:E{row-1}", '"Pending,Approved,Rejected,Deferred"')

    set_column_widths(ws, {"A": 25, "B": 22, "C": 18, "D": 14, "E": 14, "F": 28})
    ws.freeze_panes = "A4"
    logger.info("Created Approval_SignOff sheet")


def main():
    logger.info("=" * 70)
    logger.info(f"{DOCUMENT_ID} {WORKBOOK_NAME} Generator")
    logger.info("=" * 70)

    wb = Workbook()

    create_instructions_sheet(wb)
    create_executive_summary_sheet(wb)
    create_compliance_kpis_sheet(wb)
    create_authentication_events_sheet(wb)
    create_audit_findings_sheet(wb)
    create_remediation_tracker_sheet(wb)
    create_evidence_register_sheet(wb)
    create_approval_signoff_sheet(wb)

    output_path = Path(__file__).parent / OUTPUT_FILENAME
    wb.save(output_path)

    logger.info("=" * 70)
    logger.info(f"SUCCESS: Workbook saved as {OUTPUT_FILENAME}")
    logger.info("=" * 70)

    return 0


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.17 Authentication Information control
# =============================================================================
