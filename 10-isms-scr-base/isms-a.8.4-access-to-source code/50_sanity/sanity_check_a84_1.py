#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
Excel Workbook Sanity Checker - ISMS A.8.4.1 Repository Access Assessment
================================================================================

Domain-specific diagnostic utility for A.8.4.1 Repository Access Control
assessment workbooks.

**Purpose:**
Validates workbook structure, formulas, and data validations specific to
repository access control assessment requirements.

**Usage:**
    python3 sanity_check_a84_1.py ISMS_A_8_4_1_Repository_Access_YYYYMMDD.xlsx

**Checks Performed:**
- Sheet structure (10 expected sheets)
- Repository inventory completeness
- Access control validation dropdowns
- Git platform-specific validations
- Access level hierarchy
- Evidence register linkage
- Formula integrity in calculated fields

**Exit Codes:**
    0 = All checks passed
    1 = Warnings detected
    2 = Critical errors detected

Control Reference: ISO/IEC 27001:2022 Annex A Control A.8.4
Assessment Domain: 1 of 3 (Repository Access Control)
Version: 1.0
================================================================================
"""

import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def check_a84_1_workbook(filename):
    """Domain-specific validation for A.8.4.1 Repository Access assessment."""
    
    print("=" * 80)
    print("A.8.4.1 REPOSITORY ACCESS CONTROL ASSESSMENT - SANITY CHECK")
    print("=" * 80)
    print(f"File: {filename}\n")
    
    issues = []
    warnings = []
    
    # Load workbook
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"✓ Workbook loaded: {len(wb.sheetnames)} sheets\n")
    except Exception as e:
        print(f"✗ CRITICAL: Cannot load workbook: {e}")
        return 2
    
    # ========================================================================
    # EXPECTED SHEET STRUCTURE
    # ========================================================================
    print("=" * 80)
    print("CHECK 1: SHEET STRUCTURE")
    print("=" * 80)
    
    expected_sheets = [
        "Instructions & Legend",
        "Repository Inventory",
        "Access Control Assessment",
        "Access Review Records",
        "Third-Party Access",
        "Orphaned Access Analysis",
        "Audit Logging Assessment",
        "Gap Analysis",
        "Evidence Register",
        "Approval & Sign-Off"
    ]
    
    found_sheets = wb.sheetnames
    missing_sheets = [s for s in expected_sheets if s not in found_sheets]
    extra_sheets = [s for s in found_sheets if s not in expected_sheets]
    
    if missing_sheets:
        for sheet in missing_sheets:
            issues.append(f"  ✗ Missing required sheet: '{sheet}'")
    
    if extra_sheets:
        for sheet in extra_sheets:
            warnings.append(f"  ⚠ Unexpected sheet found: '{sheet}'")
    
    if not missing_sheets and not extra_sheets:
        print("  ✓ All expected sheets present, no extras")
    
    print(f"  Expected: {len(expected_sheets)} sheets")
    print(f"  Found: {len(found_sheets)} sheets")
    
    # ========================================================================
    # REPOSITORY INVENTORY VALIDATION
    # ========================================================================
    if "Repository Inventory" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 2: REPOSITORY INVENTORY DATA VALIDATION")
        print("=" * 80)
        
        ws = wb["Repository Inventory"]
        
        # Expected data validation dropdowns
        expected_validations = {
            'Git_Platform': ['GitHub', 'GitLab', 'Bitbucket', 'Azure DevOps', 
                            'AWS CodeCommit', 'Self-Hosted', 'Other'],
            'Classification': ['Production', 'Tools', 'Internal', 'Open Source', 
                              'Archived', 'Deprecated'],
            'Access_Required': ['Yes', 'No'],
            'Compliance_Status': ['Compliant', 'Non-Compliant', 'Under Review', 
                                 'Not Assessed']
        }
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
            
            # Check for key validation types
            validation_found = {key: False for key in expected_validations.keys()}
            
            for dv in ws.data_validations.dataValidation:
                if dv.formula1:
                    formula = str(dv.formula1)
                    # Check if formula contains expected values
                    for key, values in expected_validations.items():
                        if any(v in formula for v in values):
                            validation_found[key] = True
            
            for key, found in validation_found.items():
                if found:
                    print(f"  ✓ {key} validation detected")
                else:
                    warnings.append(f"  ⚠ {key} validation not detected")
        else:
            warnings.append("  ⚠ No data validations found in Repository Inventory")
    
    # ========================================================================
    # ACCESS CONTROL ASSESSMENT VALIDATION
    # ========================================================================
    if "Access Control Assessment" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 3: ACCESS CONTROL ASSESSMENT DATA VALIDATION")
        print("=" * 80)
        
        ws = wb["Access Control Assessment"]
        
        expected_access_levels = ['Read', 'Write', 'Admin', 'Owner', 'No Access']
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
            
            access_level_found = False
            for dv in ws.data_validations.dataValidation:
                if dv.formula1:
                    formula = str(dv.formula1)
                    if any(level in formula for level in expected_access_levels):
                        access_level_found = True
                        print("  ✓ Access level validation detected")
                        break
            
            if not access_level_found:
                warnings.append("  ⚠ Access level validation not detected")
        else:
            warnings.append("  ⚠ No data validations in Access Control Assessment")
    
    # ========================================================================
    # ACCESS REVIEW RECORDS VALIDATION
    # ========================================================================
    if "Access Review Records" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 4: ACCESS REVIEW RECORDS DATA VALIDATION")
        print("=" * 80)
        
        ws = wb["Access Review Records"]
        
        expected_frequencies = ['Monthly', 'Quarterly', 'Semi-Annual', 'Annual']
        expected_outcomes = ['Approved', 'Revoked', 'Modified', 'Escalated']
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
            
            freq_found = False
            outcome_found = False
            
            for dv in ws.data_validations.dataValidation:
                if dv.formula1:
                    formula = str(dv.formula1)
                    if any(f in formula for f in expected_frequencies):
                        freq_found = True
                    if any(o in formula for o in expected_outcomes):
                        outcome_found = True
            
            if freq_found:
                print("  ✓ Review frequency validation detected")
            else:
                warnings.append("  ⚠ Review frequency validation not detected")
            
            if outcome_found:
                print("  ✓ Review outcome validation detected")
            else:
                warnings.append("  ⚠ Review outcome validation not detected")
        else:
            warnings.append("  ⚠ No data validations in Access Review Records")
    
    # ========================================================================
    # THIRD-PARTY ACCESS VALIDATION
    # ========================================================================
    if "Third-Party Access" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 5: THIRD-PARTY ACCESS DATA VALIDATION")
        print("=" * 80)
        
        ws = wb["Third-Party Access"]
        
        expected_types = ['Contractor', 'Consultant', 'Partner', 'Vendor', 'Auditor']
        expected_time_bound = ['Yes', 'No']
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
            
            type_found = False
            time_bound_found = False
            
            for dv in ws.data_validations.dataValidation:
                if dv.formula1:
                    formula = str(dv.formula1)
                    if any(t in formula for t in expected_types):
                        type_found = True
                    if any(tb in formula for tb in expected_time_bound):
                        time_bound_found = True
            
            if type_found:
                print("  ✓ Third-party type validation detected")
            else:
                warnings.append("  ⚠ Third-party type validation not detected")
            
            if time_bound_found:
                print("  ✓ Time-bound access validation detected")
            else:
                warnings.append("  ⚠ Time-bound access validation not detected")
        else:
            warnings.append("  ⚠ No data validations in Third-Party Access")
    
    # ========================================================================
    # FORMULA INTEGRITY CHECK
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 6: FORMULA INTEGRITY")
    print("=" * 80)
    
    formula_count = 0
    formula_errors = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
                    formula = cell.value
                    
                    # Check for common errors
                    if formula.count('(') != formula.count(')'):
                        issues.append(f"  ✗ {sheet_name}!{cell.coordinate}: Unbalanced parentheses")
                        formula_errors += 1
                    
                    if formula.count('"') % 2 != 0:
                        issues.append(f"  ✗ {sheet_name}!{cell.coordinate}: Unbalanced quotes")
                        formula_errors += 1
    
    print(f"  Total formulas found: {formula_count}")
    if formula_errors == 0:
        print("  ✓ No formula syntax errors detected")
    else:
        print(f"  ✗ Formula errors found: {formula_errors}")
    
    # ========================================================================
    # EVIDENCE REGISTER LINKAGE
    # ========================================================================
    if "Evidence Register" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 7: EVIDENCE REGISTER STRUCTURE")
        print("=" * 80)
        
        ws = wb["Evidence Register"]
        
        # Check for expected columns (simplified check)
        has_content = False
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=10):
            for cell in row:
                if cell.value:
                    has_content = True
                    break
            if has_content:
                break
        
        if has_content:
            print("  ✓ Evidence Register has content")
        else:
            warnings.append("  ⚠ Evidence Register appears empty")
        
        # Check for data validations
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
        else:
            warnings.append("  ⚠ No data validations in Evidence Register")
    
    # ========================================================================
    # GAP ANALYSIS STRUCTURE
    # ========================================================================
    if "Gap Analysis" in wb.sheetnames:
        print("\n" + "=" * 80)
        print("CHECK 8: GAP ANALYSIS STRUCTURE")
        print("=" * 80)
        
        ws = wb["Gap Analysis"]
        
        expected_severities = ['Critical', 'High', 'Medium', 'Low']
        expected_statuses = ['Open', 'In Progress', 'Resolved', 'Accepted']
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            print(f"  ✓ Data validations present: {len(ws.data_validations.dataValidation)}")
            
            severity_found = False
            status_found = False
            
            for dv in ws.data_validations.dataValidation:
                if dv.formula1:
                    formula = str(dv.formula1)
                    if any(s in formula for s in expected_severities):
                        severity_found = True
                    if any(s in formula for s in expected_statuses):
                        status_found = True
            
            if severity_found:
                print("  ✓ Severity validation detected")
            else:
                warnings.append("  ⚠ Severity validation not detected")
            
            if status_found:
                print("  ✓ Status validation detected")
            else:
                warnings.append("  ⚠ Status validation not detected")
        else:
            warnings.append("  ⚠ No data validations in Gap Analysis")
    
    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    print("\n" + "=" * 80)
    print("VALIDATION SUMMARY")
    print("=" * 80)
    
    if issues:
        print(f"\n❌ CRITICAL ISSUES: {len(issues)}")
        for issue in issues:
            print(issue)
    
    if warnings:
        print(f"\n⚠️  WARNINGS: {len(warnings)}")
        for warning in warnings:
            print(warning)
    
    if not issues and not warnings:
        print("\n✅ ALL CHECKS PASSED")
        print("\nThe A.8.4.1 Repository Access assessment workbook appears healthy.")
        print("Ready for data collection and use.")
    else:
        print("\n" + "=" * 80)
        print("RECOMMENDATIONS:")
        print("=" * 80)
        
        if missing_sheets:
            print("\n• Regenerate workbook with generate_a84_1_repository_access.py")
            print("  to ensure all required sheets are present")
        
        if warnings:
            print("\n• Review warnings - most are informational")
            print("• Data validations improve data quality but are optional")
            print("• Ensure dropdown lists match your Git platform environment")
    
    print("\n" + "=" * 80)
    
    # Return exit code
    if issues:
        return 2
    elif warnings:
        return 1
    else:
        return 0


def main():
    if len(sys.argv) < 2:
        print("=" * 80)
        print("A.8.4.1 REPOSITORY ACCESS ASSESSMENT - SANITY CHECK")
        print("=" * 80)
        print("\nUsage: python3 sanity_check_a84_1.py <filename.xlsx>")
        print("\nExample:")
        print("  python3 sanity_check_a84_1.py ISMS_A_8_4_1_Repository_Access_20250125.xlsx")
        print("\nExit codes:")
        print("  0 = All checks passed")
        print("  1 = Warnings detected (workbook usable)")
        print("  2 = Critical errors detected (regenerate recommended)")
        print("\n" + "=" * 80)
        sys.exit(1)
    
    filename = sys.argv[1]
    exit_code = check_a84_1_workbook(filename)
    sys.exit(exit_code)


if __name__ == "__main__":
    main()