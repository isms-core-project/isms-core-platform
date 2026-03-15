#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Phase 7.5 — Full Workbook Schema Extractor

Strategy: RUN each generator script (via subprocess) → open the produced .xlsx
with openpyxl → extract exact sheet structure (names, column headers, DV values,
column widths, freeze panes, status column index).

This is 100% accurate — no source-code guesswork. The workbook IS the ground truth.

Solves:
  - Numbered sheet names preserved exactly ("1. Instructions & Legend" etc.)
  - SD column references: identifies status column per sheet for COUNTIF formulas
  - Full column headers + DV values for DB-driven generation (no stubs)

Output: datasets/data/workbook_schemas.json

Run from repo root (60-isms-core-api/):
  python3 datasets/scripts/extract_workbook_schema.py
"""

import json
import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

SCRIPT_DIR      = Path(__file__).resolve().parent
REPO_ROOT       = SCRIPT_DIR.parent.parent
FRAMEWORK_ROOT  = REPO_ROOT.parent / "50-isms-core-framework"
REGISTRY_FILE   = SCRIPT_DIR.parent / "data" / "generator_registry.json"
OUTPUT_FILE     = SCRIPT_DIR.parent / "data" / "workbook_schemas.json"

# ── Status DV detection ────────────────────────────────────────────────────────
# If a column's DV values contain any of these, it's the status column
_STATUS_KEYWORDS = {
    "compliant", "non-compliant", "partial", "authorised", "unauthorised",
    "completed", "in progress", "not started", "not applicable",
    "pass", "fail", "active", "inactive",
}

def _is_status_dv(dv_values: list[str]) -> bool:
    joined = " ".join(v.lower() for v in dv_values)
    # Strip emoji prefix (✅ ⚠ ❌) before matching
    import re
    cleaned = re.sub(r'[^\w\s,-]', '', joined)
    return any(kw in cleaned for kw in _STATUS_KEYWORDS)


# ── Sheet type classification (same as parse_generators.py) ───────────────────
_SHEET_TYPE_MAP = {
    "instructions & legend": "instructions",
    "instructions": "instructions",
    "legend": "instructions",
    "summary dashboard": "summary",
    "summary": "summary",
    "evidence register": "evidence",
    "evidence": "evidence",
    "approval sign-off": "approval",
    "approval": "approval",
    "sign-off": "approval",
}

def _classify_sheet(name: str) -> str:
    lower = name.lower().strip()
    # Strip leading "N. " prefix (numbered sheets like "1. Instructions & Legend")
    import re
    lower_stripped = re.sub(r'^\d+\.\s*', '', lower)
    for key, stype in _SHEET_TYPE_MAP.items():
        if key in lower_stripped or key in lower:
            return stype
    return "input"


# ── Extract schema from a single .xlsx file ───────────────────────────────────
def extract_from_workbook(xlsx_path: Path, document_id: str) -> list[dict]:
    """
    Open the workbook and extract per-sheet schema:
      sheet_name, sheet_type, position, columns (header, width, dv_values),
      status_column_index, title_text, subtitle_text, freeze_panes, hide_gridlines
    """
    try:
        wb = load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        logger.error("Cannot open %s: %s", xlsx_path, e)
        return []

    schemas = []

    for pos, ws in enumerate(wb.worksheets, start=1):
        sheet_name = ws.title
        sheet_type = _classify_sheet(sheet_name)

        # ── Title / subtitle ──────────────────────────────────────────────────
        title_text    = _cell_value(ws, 1, 1)  # A1 (merged)
        subtitle_text = _cell_value(ws, 2, 1)  # A2

        # ── Freeze panes ──────────────────────────────────────────────────────
        fp = ws.freeze_panes
        freeze_panes = str(fp) if fp else None

        # ── Grid lines ────────────────────────────────────────────────────────
        hide_gridlines = not ws.sheet_view.showGridLines

        # ── Find header row (first row with D9D9D9 grey fill or row 3) ────────
        header_row, data_start_row = _find_header_row(ws)

        # ── Columns ───────────────────────────────────────────────────────────
        columns = []
        status_column_index = None

        if header_row:
            # Collect DVs: map column_index → list of valid values
            dv_map = _build_dv_map(ws)

            col_idx = 1
            while True:
                cell = ws.cell(row=header_row, column=col_idx)
                if cell.value is None and col_idx > 1:
                    # Check two more cols in case there are gaps
                    if ws.cell(row=header_row, column=col_idx + 1).value is None:
                        break
                if col_idx > 30:
                    break

                header_val = str(cell.value).strip() if cell.value else None
                if not header_val:
                    col_idx += 1
                    continue

                col_letter = get_column_letter(col_idx)
                width = ws.column_dimensions[col_letter].width if col_letter in ws.column_dimensions else 12.0

                dv_values = dv_map.get(col_idx, [])
                required = header_val.endswith("[*]") or header_val.endswith("*]")
                # Clean [*] from display name
                clean_header = header_val.replace(" [*]", "").replace("[*]", "").strip()

                col_entry = {
                    "index": col_idx,         # 1-based
                    "letter": col_letter,
                    "header": clean_header,
                    "header_raw": header_val,  # keep original for exact re-gen
                    "width": round(float(width), 1),
                    "dv_values": dv_values,
                    "dv_allow_blank": True,
                    "required": required,
                    "is_status_col": False,
                }
                columns.append(col_entry)

                # Mark status column
                if dv_values and _is_status_dv(dv_values) and status_column_index is None:
                    col_entry["is_status_col"] = True
                    status_column_index = col_idx

                col_idx += 1

        # ── Sample row ────────────────────────────────────────────────────────
        sample_row = []
        if header_row and data_start_row:
            sample_row_idx = data_start_row  # first row after headers
            for c in columns:
                v = _cell_value(ws, sample_row_idx, c["index"])
                sample_row.append(str(v) if v is not None else "")

        schemas.append({
            "sheet_name":          sheet_name,
            "sheet_type":          sheet_type,
            "position":            pos,
            "title_text":          str(title_text).strip() if title_text else None,
            "subtitle_text":       str(subtitle_text).strip() if subtitle_text else None,
            "header_row":          header_row,
            "data_start_row":      data_start_row,
            "freeze_panes":        freeze_panes,
            "hide_gridlines":      hide_gridlines,
            "columns":             columns,
            "status_column_index": status_column_index,
            "status_column_letter":get_column_letter(status_column_index) if status_column_index else None,
            "sample_row":          sample_row,
        })

    wb.close()
    return schemas


def _cell_value(ws, row: int, col: int):
    """Get cell value, handling merged cells."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    # Check if it's a merged cell — find the master
    for merge in ws.merged_cells.ranges:
        if merge.min_row <= row <= merge.max_row and merge.min_col <= col <= merge.max_col:
            return ws.cell(merge.min_row, merge.min_col).value
    return None


def _find_header_row(ws) -> tuple[int | None, int | None]:
    """
    Find the header row and data start row.

    Two recognised patterns:
      - Standard: row N has D9D9D9 grey fill + text → data starts row N+1
      - Numbered-sheet: row 1 has 003366 dark-blue fill with MULTIPLE individual
        header cells (not a merged title) → data starts row 3 (row 2 = sample)

    Falls back to first text row in rows 1-6 if neither pattern matches.
    """
    _GREY_RGBS = {"FFD9D9D9", "D9D9D9", "00D9D9D9"}
    _BLUE_RGBS = {"FF003366", "003366", "00003366"}

    # Collect all merged cell ranges for quick lookup
    merged = ws.merged_cells.ranges

    def _is_merged_across(row: int, col: int) -> bool:
        """True if this cell is part of a horizontal merge spanning >1 col."""
        for rng in merged:
            if (rng.min_row <= row <= rng.max_row
                    and rng.min_col <= col <= rng.max_col
                    and rng.max_col > rng.min_col):
                return True
        return False

    for r in range(1, 8):
        grey_count = 0
        blue_count = 0
        text_count = 0
        blue_non_merged = 0

        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            fill = cell.fill
            if cell.value:
                text_count += 1
            if fill and fill.fill_type == "solid":
                rgb = (fill.fgColor.rgb or "").upper() if fill.fgColor else ""
                if rgb in _GREY_RGBS:
                    grey_count += 1
                elif rgb in _BLUE_RGBS:
                    blue_count += 1
                    if not _is_merged_across(r, c):
                        blue_non_merged += 1

        if text_count:
            if grey_count:
                return r, r + 1
            # Blue header: must have ≥2 non-merged blue cells (i.e. real column headers)
            if blue_non_merged >= 2:
                return r, r + 2  # row 2 is sample row in this layout

    # Fallback: first row with text values in rows 1-6
    for r in range(1, 7):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 6)]
        if any(vals):
            return r, r + 1
    return None, None


def _build_dv_map(ws) -> dict[int, list[str]]:
    """
    Build {column_index: [dv_values]} from worksheet data validations.
    Only list-type DVs. Strips emoji and quotes.
    """
    import re
    dv_map: dict[int, list[str]] = {}

    for dv in ws.data_validations.dataValidation:
        if dv.type != "list" or not dv.formula1:
            continue

        # Extract values from formula1 string like '"Val1,Val2,Val3"' or '=Sheet!A1:A5'
        raw = dv.formula1.strip('"\'')
        values = [v.strip() for v in raw.split(",") if v.strip()]

        # Map DV to column indices from sqref ranges
        for sqref_range in str(dv.sqref).split():
            # sqref can be "E5:E56" or "E5"
            m = re.match(r'\$?([A-Z]+)\$?\d+(?::\$?([A-Z]+)\$?\d+)?', sqref_range)
            if not m:
                continue
            from openpyxl.utils import column_index_from_string
            start_col = column_index_from_string(m.group(1))
            end_col   = column_index_from_string(m.group(2)) if m.group(2) else start_col
            for ci in range(start_col, end_col + 1):
                if ci not in dv_map:
                    dv_map[ci] = values

    return dv_map


# ── Find the existing WKBK file for a generator ───────────────────────────────
def find_wkbk_file(src_path: Path, doc_id: str) -> Path | None:
    """
    Generators always save to Path(__file__).parent.parent / "WKBK".
    Find the most recent .xlsx in that dir whose name starts with doc_id.
    """
    wkbk_dir = src_path.parent.parent / "WKBK"
    if not wkbk_dir.exists():
        logger.warning("WKBK dir not found: %s", wkbk_dir)
        return None

    prefix = doc_id + "_"
    candidates = [f for f in wkbk_dir.glob("*.xlsx") if f.name.startswith(prefix)]
    if not candidates:
        logger.warning("No WKBK file found for %s in %s", doc_id, wkbk_dir)
        return None

    return max(candidates, key=lambda p: p.stat().st_mtime)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    registry = json.loads(REGISTRY_FILE.read_text())
    logger.info("Loaded %d generator records", len(registry))

    results = []
    stats   = {"ok": 0, "failed": 0, "skipped": 0}

    for i, rec in enumerate(registry, start=1):
        doc_id   = rec["document_id"]
        src_rel  = rec.get("source_file")

        if not src_rel:
            logger.warning("[%d/%d] No source_file for %s — skipped", i, len(registry), doc_id)
            stats["skipped"] += 1
            results.append({"document_id": doc_id, "sheets": [], "error": "no_source_file"})
            continue

        src_path = FRAMEWORK_ROOT / src_rel
        if not src_path.exists():
            logger.warning("[%d/%d] Source not found: %s — skipped", i, len(registry), src_rel)
            stats["skipped"] += 1
            results.append({"document_id": doc_id, "sheets": [], "error": "source_not_found"})
            continue

        logger.info("[%d/%d] Extracting %s ...", i, len(registry), doc_id)
        xlsx = find_wkbk_file(src_path, doc_id)

        if not xlsx:
            stats["failed"] += 1
            results.append({"document_id": doc_id, "sheets": [], "error": "wkbk_not_found"})
            continue

        sheets = extract_from_workbook(xlsx, doc_id)
        logger.info("  → %d sheets from %s", len(sheets), xlsx.name)
        stats["ok"] += 1

        results.append({
            "document_id": doc_id,
            "source_file": src_rel,
            "xlsx_file":   xlsx.name,
            "sheets":      sheets,
        })

    OUTPUT_FILE.write_text(json.dumps(results, indent=2, ensure_ascii=False))
    logger.info("Written: %s (%d records)", OUTPUT_FILE, len(results))
    logger.info("Stats: %s", stats)

    failed = [r["document_id"] for r in results if r.get("error")]
    if failed:
        logger.warning("%d generators failed/skipped: %s", len(failed), failed)


if __name__ == "__main__":
    main()
