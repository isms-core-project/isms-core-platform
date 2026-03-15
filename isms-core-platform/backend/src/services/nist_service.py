"""Service layer for NIST CSF 2.0 assessment profiles."""

import csv
import io
import logging
import re
import uuid
from datetime import datetime, timezone

from sqlalchemy import func, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from src.domain.frameworks import CrossFrameworkMapping, Framework, FrameworkControl
from src.domain.nist import NistCsfProfile, NistCsfRating
from src.schemas.nist import (
    FunctionScore,
    NistFullProfile,
    NistProfileCreate,
    NistProfileRead,
    NistProfileSummary,
    NistProfileUpdate,
    NistRatingRead,
    NistRatingUpsert,
    NistSubcategory,
)

logger = logging.getLogger(__name__)

NIST_FRAMEWORK_CODE = "NIST_CSF_2.0"
TOTAL_SUBCATEGORIES = 106

TIER_LABELS = {
    1: "Partial",
    2: "Risk Informed",
    3: "Repeatable",
    4: "Adaptive",
}

FUNCTION_NAMES = {
    "GV": "Govern",
    "ID": "Identify",
    "PR": "Protect",
    "DE": "Detect",
    "RS": "Respond",
    "RC": "Recover",
}


# ── Framework helpers ─────────────────────────────────────────────────────────

def _get_nist_framework(db: Session) -> Framework:
    fw = db.execute(
        select(Framework).where(Framework.code == NIST_FRAMEWORK_CODE)
    ).scalar_one_or_none()
    if not fw:
        raise ValueError(
            "NIST CSF 2.0 framework not loaded — run /admin/load first."
        )
    return fw


def _get_subcategories(db: Session, framework_id: uuid.UUID) -> list[FrameworkControl]:
    return list(
        db.execute(
            select(FrameworkControl)
            .where(
                FrameworkControl.framework_id == framework_id,
                FrameworkControl.level == 2,
            )
            .order_by(FrameworkControl.sort_order)
        ).scalars().all()
    )


def _build_iso_mapping(db: Session, subcategory_ids: list[uuid.UUID]) -> dict[uuid.UUID, list[str]]:
    """Return {nist_subcategory_id: [iso_control_id, ...]} for all given subcategory UUIDs.

    Crosswalk direction in DB: ISO27001_2022 (source) → NIST_CSF_2.0 (target).
    So NIST subcategory_ids are in target_control_id.
    """
    if not subcategory_ids:
        return {}

    rows = db.execute(
        select(
            CrossFrameworkMapping.target_control_id,   # NIST subcategory
            FrameworkControl.control_id,               # ISO control e.g. A.5.1
        )
        .join(
            FrameworkControl,
            FrameworkControl.id == CrossFrameworkMapping.source_control_id,
        )
        .join(
            Framework,
            Framework.id == FrameworkControl.framework_id,
        )
        .where(
            CrossFrameworkMapping.target_control_id.in_(subcategory_ids),
            Framework.code == "ISO27001_2022",
        )
    ).all()

    result: dict[uuid.UUID, list[str]] = {}
    for nist_id, iso_code in rows:
        result.setdefault(nist_id, []).append(iso_code)
    return result


# ── Profile CRUD ──────────────────────────────────────────────────────────────

def list_profiles(db: Session) -> list[NistProfileSummary]:
    profiles = list(
        db.execute(
            select(NistCsfProfile).order_by(NistCsfProfile.created_at.desc())
        ).scalars().all()
    )
    summaries = []
    for p in profiles:
        summary = _build_profile_summary(db, p)
        summaries.append(summary)
    return summaries


def create_profile(db: Session, data: NistProfileCreate) -> NistProfileRead:
    profile = NistCsfProfile(
        name=data.name,
        description=data.description,
        assessor=data.assessor,
        scope=data.scope,
        status="draft",
    )
    db.add(profile)
    db.commit()
    db.refresh(profile)
    logger.info("NIST CSF profile created: %s (%s)", profile.name, profile.id)
    return NistProfileRead.model_validate(profile)


def update_profile(db: Session, profile_id: uuid.UUID, data: NistProfileUpdate) -> NistProfileRead:
    profile = db.get(NistCsfProfile, profile_id)
    if not profile:
        raise ValueError(f"Profile {profile_id} not found.")
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(profile, field, value)
    db.commit()
    db.refresh(profile)
    return NistProfileRead.model_validate(profile)


def delete_profile(db: Session, profile_id: uuid.UUID) -> bool:
    profile = db.get(NistCsfProfile, profile_id)
    if not profile:
        return False
    db.delete(profile)
    db.commit()
    logger.info("NIST CSF profile deleted: %s", profile_id)
    return True


# ── Full profile ──────────────────────────────────────────────────────────────

def get_full_profile(db: Session, profile_id: uuid.UUID) -> NistFullProfile:
    profile = db.get(NistCsfProfile, profile_id)
    if not profile:
        raise ValueError(f"Profile {profile_id} not found.")

    fw = _get_nist_framework(db)
    subcategories = _get_subcategories(db, fw.id)
    sub_ids = [s.id for s in subcategories]

    # Existing ratings keyed by subcategory_id
    existing: dict[uuid.UUID, NistCsfRating] = {
        r.subcategory_id: r
        for r in db.execute(
            select(NistCsfRating).where(NistCsfRating.profile_id == profile_id)
        ).scalars().all()
    }

    iso_map = _build_iso_mapping(db, sub_ids)

    ratings: list[NistRatingRead] = []
    for sub in subcategories:
        meta = sub.metadata_ or {}
        rating = existing.get(sub.id)
        ratings.append(NistRatingRead(
            id=rating.id if rating else uuid.uuid4(),
            profile_id=profile_id,
            subcategory_id=sub.id,
            subcategory_code=sub.control_id,
            subcategory_title=sub.title,
            function_code=meta.get("function", ""),
            category_code=meta.get("category", ""),
            current_tier=rating.current_tier if rating else None,
            target_tier=rating.target_tier if rating else None,
            notes=rating.notes if rating else None,
            iso_mappings=iso_map.get(sub.id, []),
        ))

    return NistFullProfile(
        profile=NistProfileRead.model_validate(profile),
        ratings=ratings,
    )


# ── Summary ───────────────────────────────────────────────────────────────────

def _build_profile_summary(db: Session, profile: NistCsfProfile) -> NistProfileSummary:
    rows = db.execute(
        select(
            NistCsfRating.current_tier,
            NistCsfRating.target_tier,
            FrameworkControl.metadata_["function"].astext.label("function_code"),
        )
        .join(FrameworkControl, FrameworkControl.id == NistCsfRating.subcategory_id)
        .where(NistCsfRating.profile_id == profile.id)
    ).all()

    rated = [r for r in rows if r.current_tier is not None]
    all_current = [r.current_tier for r in rated]
    all_target = [r.target_tier for r in rows if r.target_tier is not None]

    # Per-function aggregates
    func_data: dict[str, dict] = {}
    for fn_code in FUNCTION_NAMES:
        func_data[fn_code] = {"current": [], "target": [], "rated": 0}
    for r in rows:
        fc = r.function_code or ""
        if fc in func_data:
            if r.current_tier is not None:
                func_data[fc]["current"].append(r.current_tier)
                func_data[fc]["rated"] += 1
            if r.target_tier is not None:
                func_data[fc]["target"].append(r.target_tier)

    function_scores = [
        FunctionScore(
            function_code=fc,
            function_name=FUNCTION_NAMES.get(fc, fc),
            avg_current=round(sum(d["current"]) / len(d["current"]), 2) if d["current"] else None,
            avg_target=round(sum(d["target"]) / len(d["target"]), 2) if d["target"] else None,
            rated_count=d["rated"],
            total_count=_FUNCTION_COUNTS.get(fc, 0),
        )
        for fc, d in func_data.items()
    ]

    return NistProfileSummary(
        **NistProfileRead.model_validate(profile).model_dump(),
        rated_count=len(rated),
        total_subcategories=TOTAL_SUBCATEGORIES,
        avg_current_tier=round(sum(all_current) / len(all_current), 2) if all_current else None,
        avg_target_tier=round(sum(all_target) / len(all_target), 2) if all_target else None,
        function_scores=function_scores,
    )


def get_profile_summary(db: Session, profile_id: uuid.UUID) -> NistProfileSummary:
    profile = db.get(NistCsfProfile, profile_id)
    if not profile:
        raise ValueError(f"Profile {profile_id} not found.")
    return _build_profile_summary(db, profile)


# NIST CSF 2.0 subcategory counts per function (fixed values)
_FUNCTION_COUNTS = {"GV": 23, "ID": 5, "PR": 20, "DE": 9, "RS": 17, "RC": 6}


# ── Rating upsert ─────────────────────────────────────────────────────────────

def upsert_ratings(
    db: Session, profile_id: uuid.UUID, ratings: list[NistRatingUpsert]
) -> list[NistRatingRead]:
    profile = db.get(NistCsfProfile, profile_id)
    if not profile:
        raise ValueError(f"Profile {profile_id} not found.")

    now = datetime.now(timezone.utc)

    for r in ratings:
        stmt = (
            pg_insert(NistCsfRating)
            .values(
                id=uuid.uuid4(),
                profile_id=profile_id,
                subcategory_id=r.subcategory_id,
                current_tier=r.current_tier,
                target_tier=r.target_tier,
                notes=r.notes,
                updated_at=now,
            )
            .on_conflict_do_update(
                constraint="uq_nist_rating_profile_subcategory",
                set_={
                    "current_tier": r.current_tier,
                    "target_tier": r.target_tier,
                    "notes": r.notes,
                    "updated_at": now,
                },
            )
        )
        db.execute(stmt)

    # Touch profile updated_at
    profile.updated_at = now
    db.commit()

    logger.info("Upserted %d NIST ratings for profile %s", len(ratings), profile_id)

    # Return updated ratings for modified subcategories
    sub_ids = [r.subcategory_id for r in ratings]
    updated_rows = db.execute(
        select(NistCsfRating, FrameworkControl)
        .join(FrameworkControl, FrameworkControl.id == NistCsfRating.subcategory_id)
        .where(
            NistCsfRating.profile_id == profile_id,
            NistCsfRating.subcategory_id.in_(sub_ids),
        )
    ).all()

    iso_map = _build_iso_mapping(db, sub_ids)

    result = []
    for rating, sub in updated_rows:
        meta = sub.metadata_ or {}
        result.append(NistRatingRead(
            id=rating.id,
            profile_id=rating.profile_id,
            subcategory_id=rating.subcategory_id,
            subcategory_code=sub.control_id,
            subcategory_title=sub.title,
            function_code=meta.get("function", ""),
            category_code=meta.get("category", ""),
            current_tier=rating.current_tier,
            target_tier=rating.target_tier,
            notes=rating.notes,
            iso_mappings=iso_map.get(sub.id, []),
        ))
    return result


# ── Subcategories (seed for UI grid) ─────────────────────────────────────────

def list_subcategories(db: Session) -> list[NistSubcategory]:
    fw = _get_nist_framework(db)
    subcategories = _get_subcategories(db, fw.id)
    sub_ids = [s.id for s in subcategories]
    iso_map = _build_iso_mapping(db, sub_ids)

    return [
        NistSubcategory(
            id=sub.id,
            control_id=sub.control_id,
            title=sub.title,
            function_code=(sub.metadata_ or {}).get("function", ""),
            category_code=(sub.metadata_ or {}).get("category", ""),
            sort_order=sub.sort_order,
            iso_mappings=iso_map.get(sub.id, []),
        )
        for sub in subcategories
    ]


# ── CSV export ────────────────────────────────────────────────────────────────

def export_profile_csv(db: Session, profile_id: uuid.UUID) -> str:
    full = get_full_profile(db, profile_id)
    profile_name = full.profile.name

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Function", "Category", "Subcategory", "Title",
        "Current Tier", "Current Tier Label",
        "Target Tier", "Target Tier Label",
        "Gap", "Notes", "ISO 27001 Mappings",
    ])

    for r in full.ratings:
        gap = ""
        if r.current_tier is not None and r.target_tier is not None:
            gap = str(r.target_tier - r.current_tier)
        writer.writerow([
            r.function_code,
            r.category_code,
            r.subcategory_code,
            r.subcategory_title,
            r.current_tier or "",
            TIER_LABELS.get(r.current_tier, "") if r.current_tier else "",
            r.target_tier or "",
            TIER_LABELS.get(r.target_tier, "") if r.target_tier else "",
            gap,
            r.notes or "",
            "; ".join(r.iso_mappings),
        ])

    logger.info("CSV export generated for NIST profile '%s'", profile_name)
    return output.getvalue()


# ── XLSX import ───────────────────────────────────────────────────────────────

_SUBCATEGORY_RE = re.compile(r"^([A-Z]{2}\.[A-Z]+-\d+)")


def import_from_xlsx(db: Session, profile_id: uuid.UUID, file_bytes: bytes) -> dict:
    """Parse an official NIST CSF 2.0 Organisational Profile Template XLSX and upsert ratings.

    Handles both formats:
    - Official template (sheet "Current and Target Profile", col A = code, col M = Target CSF Tier)
    - Reference tool with user-added tier columns (sheet "CSF 2.0", col C = "CODE: description")
    """
    profile = db.get(NistCsfProfile, profile_id)
    if not profile:
        raise ValueError(f"Profile {profile_id} not found.")

    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read workbook: {exc}") from exc

    # Prefer official template sheet, then reference tool sheet, then first sheet
    ws = None
    for candidate in ("Current and Target Profile", "CSF 2.0"):
        if candidate in wb.sheetnames:
            ws = wb[candidate]
            break
    if ws is None:
        ws = wb.active

    # Find header row (first row containing 'subcategory' or 'outcome')
    header_row_idx: int | None = None
    headers: list[str] = []
    for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
        row_strs = [str(v).lower() if v else "" for v in row]
        if any("subcategory" in v or "outcome" in v for v in row_strs):
            header_row_idx = row_idx
            headers = row_strs
            break

    if header_row_idx is None:
        raise ValueError(
            "Cannot find header row — expected a column containing 'Subcategory' or 'CSF Outcome'."
        )

    # Detect column indices
    code_col = next(
        (i for i, h in enumerate(headers) if "subcategory" in h or "outcome" in h), 0
    )
    current_col = next(
        (i for i, h in enumerate(headers) if "current" in h and "tier" in h), None
    )
    if current_col is None:
        # Fall back to "current priority" — used by some templates for numeric ratings
        current_col = next(
            (i for i, h in enumerate(headers) if "current" in h and "priority" in h), None
        )
    target_col = next(
        (i for i, h in enumerate(headers) if "target" in h and "tier" in h), None
    )
    notes_col = next((i for i, h in enumerate(headers) if h == "notes"), None)

    # Build subcategory lookup: control_id → ORM object
    fw = _get_nist_framework(db)
    subcategories = _get_subcategories(db, fw.id)
    code_to_sub = {sub.control_id: sub for sub in subcategories}

    def _to_tier(val) -> int | None:
        if val is None:
            return None
        try:
            n = int(float(str(val)))
            return n if 1 <= n <= 4 else None
        except (ValueError, TypeError):
            return None

    to_upsert: list[NistRatingUpsert] = []
    skipped = 0
    not_found: list[str] = []

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or not row[code_col]:
            skipped += 1
            continue
        cell_val = str(row[code_col]).strip()
        m = _SUBCATEGORY_RE.match(cell_val)
        if not m:
            skipped += 1
            continue

        code = m.group(1)
        sub = code_to_sub.get(code)
        if sub is None:
            not_found.append(code)
            continue

        def _col_val(idx: int | None):
            if idx is None or len(row) <= idx:
                return None
            return row[idx]

        current_tier = _to_tier(_col_val(current_col))
        target_tier = _to_tier(_col_val(target_col))
        notes_raw = _col_val(notes_col)
        notes = str(notes_raw).strip() or None if notes_raw else None

        if current_tier is not None or target_tier is not None or notes:
            to_upsert.append(
                NistRatingUpsert(
                    subcategory_id=sub.id,
                    current_tier=current_tier,
                    target_tier=target_tier,
                    notes=notes,
                )
            )

    if to_upsert:
        upsert_ratings(db, profile_id, to_upsert)

    logger.info(
        "NIST XLSX import: %d upserted, %d skipped, %d not found — profile %s",
        len(to_upsert), skipped, len(not_found), profile_id,
    )
    return {"imported": len(to_upsert), "skipped": skipped, "not_found": not_found}


# ── XLSX export ───────────────────────────────────────────────────────────────

def export_profile_xlsx(db: Session, profile_id: uuid.UUID) -> bytes:
    """Generate a formatted XLSX workbook with Summary, Assessment, and Gap Analysis sheets."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    full = get_full_profile(db, profile_id)
    summary = _build_profile_summary(db, db.get(NistCsfProfile, profile_id))  # type: ignore[arg-type]
    profile = full.profile

    # ── Style constants ────────────────────────────────────────────────────────
    tier_fills = {
        1: PatternFill(start_color="FFFF5252", end_color="FFFF5252", fill_type="solid"),
        2: PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"),
        3: PatternFill(start_color="FF70AD47", end_color="FF70AD47", fill_type="solid"),
        4: PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid"),
    }
    header_fill = PatternFill(start_color="FF1F3864", end_color="FF1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF", size=10)
    bold_font = Font(bold=True, size=10)
    thin = Side(style="thin", color="FFAAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    def apply_header(ws, row: int, cols: list[str], widths: list[float] | None = None) -> None:
        for c, h in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center
        if widths:
            for c, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(c)].width = w

    def write_rating_row(ws, r_idx: int, rating) -> None:
        gap = None
        if rating.current_tier is not None and rating.target_tier is not None:
            gap = rating.target_tier - rating.current_tier
        row_vals = [
            rating.function_code,
            rating.category_code,
            rating.subcategory_code,
            rating.subcategory_title,
            rating.current_tier,
            TIER_LABELS.get(rating.current_tier, "") if rating.current_tier else "",
            rating.target_tier,
            TIER_LABELS.get(rating.target_tier, "") if rating.target_tier else "",
            gap,
            rating.notes or "",
            "; ".join(rating.iso_mappings),
        ]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if c_idx == 4:
                cell.alignment = wrap
            elif c_idx in (1, 2, 3, 5, 7, 9):
                cell.alignment = center
            if c_idx == 5 and rating.current_tier in tier_fills:
                cell.fill = tier_fills[rating.current_tier]
                cell.font = Font(bold=True, color="FFFFFFFF", size=10)
            if c_idx == 7 and rating.target_tier in tier_fills:
                cell.fill = tier_fills[rating.target_tier]
                cell.font = Font(bold=True, color="FFFFFFFF", size=10)
            if c_idx == 9 and gap and gap > 0:
                cell.fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"

    ws1.merge_cells("A1:G1")
    t = ws1["A1"]
    t.value = f"NIST CSF 2.0 Assessment — {profile.name}"
    t.font = Font(bold=True, size=14, color="FF4472C4")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 28

    meta_rows = [
        ("Assessor", profile.assessor or "—"),
        ("Scope", profile.scope or "—"),
        ("Status", profile.status.capitalize()),
        ("Date", profile.updated_at.strftime("%Y-%m-%d")),
        ("Coverage", f"{summary.rated_count} / {summary.total_subcategories} subcategories"),
        ("Avg Current Tier", f"T{summary.avg_current_tier:.1f}" if summary.avg_current_tier else "—"),
        ("Avg Target Tier", f"T{summary.avg_target_tier:.1f}" if summary.avg_target_tier else "—"),
    ]
    for i, (label, val) in enumerate(meta_rows, start=2):
        ws1.cell(row=i, column=1, value=label).font = bold_font
        ws1.cell(row=i, column=2, value=val)

    func_row = len(meta_rows) + 3
    apply_header(
        ws1, func_row,
        ["Function", "Name", "Avg Current", "Avg Target", "Gap", "Rated", "Total"],
        [10, 16, 14, 14, 8, 8, 8],
    )
    for i, fs in enumerate(summary.function_scores, start=func_row + 1):
        gap_v = None
        if fs.avg_current is not None and fs.avg_target is not None:
            gap_v = round(fs.avg_target - fs.avg_current, 2)
        row_vals = [
            fs.function_code,
            fs.function_name,
            round(fs.avg_current, 2) if fs.avg_current else "—",
            round(fs.avg_target, 2) if fs.avg_target else "—",
            gap_v if gap_v is not None else "—",
            fs.rated_count,
            fs.total_count,
        ]
        for c, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=i, column=c, value=val)
            cell.border = border
            cell.alignment = center
            if c == 3 and fs.avg_current:
                t = round(fs.avg_current)
                if t in tier_fills:
                    cell.fill = tier_fills[t]
                    cell.font = Font(bold=True, color="FFFFFFFF", size=10)

    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 16

    # ── Sheet 2: Assessment ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Assessment")
    assess_cols = [
        "Function", "Category", "Subcategory", "Title",
        "Current Tier", "Current Label", "Target Tier", "Target Label",
        "Gap", "Notes", "ISO 27001 Mappings",
    ]
    apply_header(ws2, 1, assess_cols, [10, 12, 14, 50, 13, 16, 13, 16, 6, 30, 32])
    for r_idx, rating in enumerate(full.ratings, start=2):
        write_rating_row(ws2, r_idx, rating)
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Gap Analysis ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Gap Analysis")
    apply_header(ws3, 1, assess_cols, [10, 12, 14, 50, 13, 16, 13, 16, 6, 30, 32])
    gap_ratings = sorted(
        [r for r in full.ratings if r.current_tier is not None and r.target_tier is not None
         and r.target_tier > r.current_tier],
        key=lambda r: r.target_tier - r.current_tier,  # type: ignore[operator]
        reverse=True,
    )
    for r_idx, rating in enumerate(gap_ratings, start=2):
        write_rating_row(ws3, r_idx, rating)
    ws3.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info("XLSX export generated for NIST profile '%s'", profile.name)
    return output.getvalue()
