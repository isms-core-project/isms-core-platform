import csv
import io
import uuid
from datetime import datetime, timezone

from sqlalchemy import select
from sqlalchemy.orm import Session as DBSession

from src.domain.assessments import Assessment, AssessmentItem
from src.domain.collections import AssessmentCollection, collection_assessments
from src.domain.control_groups import ControlGroup
from src.schemas.collections import CollectionCreate, CollectionMember, CollectionPatch, CollectionStats


def _compute_stats(assessments: list) -> CollectionStats:
    total = len(assessments)
    started = sum(1 for a in assessments if (a.items_total or 0) > 0)
    items_total = sum(a.items_total or 0 for a in assessments)
    items_compliant = sum(a.items_compliant or 0 for a in assessments)
    items_non_compliant = sum(a.items_non_compliant or 0 for a in assessments)
    completion_pct = round(started / total * 100, 1) if total > 0 else 0.0
    compliance_pct = round(items_compliant / items_total * 100, 1) if items_total > 0 else 0.0
    if started == 0:
        status = "not_started"
    elif started == total:
        status = "complete"
    else:
        status = "in_progress"
    return CollectionStats(
        total=total,
        started=started,
        completion_pct=completion_pct,
        items_total=items_total,
        items_compliant=items_compliant,
        items_non_compliant=items_non_compliant,
        compliance_pct=compliance_pct,
        status=status,
    )


def _build_members(db: DBSession, assessments: list) -> list[CollectionMember]:
    members = []
    for a in assessments:
        cg = db.get(ControlGroup, a.control_group_id)
        items_total = a.items_total or 0
        items_compliant = a.items_compliant or 0
        items_non_compliant = a.items_non_compliant or 0
        compliance_pct = round(items_compliant / items_total * 100, 1) if items_total > 0 else 0.0
        members.append(CollectionMember(
            assessment_id=str(a.id),
            document_id=a.document_id,
            workbook_name=a.workbook_name,
            group_code=cg.group_code.upper() if cg else "",
            group_name=cg.name if cg else "",
            product_type=a.product_type.value if hasattr(a.product_type, "value") else str(a.product_type),
            items_total=items_total,
            items_compliant=items_compliant,
            items_non_compliant=items_non_compliant,
            compliance_pct=compliance_pct,
        ))
    return members


def list_collections(db: DBSession, product_family: str | None = None) -> list[AssessmentCollection]:
    stmt = select(AssessmentCollection).order_by(AssessmentCollection.created_at.desc())
    if product_family:
        stmt = stmt.where(AssessmentCollection.product_family == product_family.upper())
    return list(db.execute(stmt).scalars().all())


def get_collection(db: DBSession, collection_id: uuid.UUID) -> AssessmentCollection | None:
    return db.get(AssessmentCollection, collection_id)


def create_collection(db: DBSession, body: CollectionCreate) -> AssessmentCollection:
    coll = AssessmentCollection(
        name=body.name,
        description=body.description,
        product_family=body.product_family.upper(),
        product_type=body.product_type,
        due_date=body.due_date,
    )
    db.add(coll)
    db.commit()
    db.refresh(coll)
    return coll


def patch_collection(db: DBSession, coll: AssessmentCollection, body: CollectionPatch) -> AssessmentCollection:
    if body.name is not None:
        coll.name = body.name
    if body.description is not None:
        coll.description = body.description
    if body.due_date is not None:
        coll.due_date = body.due_date
    db.commit()
    db.refresh(coll)
    return coll


def delete_collection(db: DBSession, coll: AssessmentCollection) -> None:
    db.delete(coll)
    db.commit()


def add_assessment(db: DBSession, coll: AssessmentCollection, assessment_id: uuid.UUID) -> None:
    assessment = db.get(Assessment, assessment_id)
    if not assessment:
        raise ValueError("Assessment not found")
    if assessment not in coll.assessments:
        coll.assessments.append(assessment)
        db.commit()


def remove_assessment(db: DBSession, coll: AssessmentCollection, assessment_id: uuid.UUID) -> None:
    assessment = db.get(Assessment, assessment_id)
    if assessment and assessment in coll.assessments:
        coll.assessments.remove(assessment)
        db.commit()


def get_non_compliant_items(db: DBSession, assessments: list) -> list[dict]:
    """Return non-compliant items across all assessments in the collection."""
    results = []
    for a in assessments:
        cg = db.get(ControlGroup, a.control_group_id)
        items = db.execute(
            select(AssessmentItem).where(
                AssessmentItem.assessment_id == a.id,
                AssessmentItem.status == "non_compliant",
            )
        ).scalars().all()
        for item in items:
            results.append({
                "assessment": a.workbook_name,
                "group_code": cg.group_code.upper() if cg else "",
                "group_name": cg.name if cg else "",
                "item_text": item.item_text or "",
                "notes": item.notes or "",
            })
    return results


# ── CSV ────────────────────────────────────────────────────────────────────────

def export_csv(db: DBSession, coll: AssessmentCollection) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)

    # Summary
    stats = _compute_stats(coll.assessments)
    writer.writerow(["Collection", coll.name])
    writer.writerow(["Product", f"{coll.product_family} / {coll.product_type or ''}".strip(" /")])
    writer.writerow(["Status", stats.status.replace("_", " ").title()])
    writer.writerow(["Completion", f"{stats.completion_pct}%"])
    writer.writerow(["Compliance Score", f"{stats.compliance_pct}%"])
    writer.writerow([])

    # Per-assessment scores
    writer.writerow(["Control Group", "Workbook", "Product Type", "Items Total", "Compliant", "Non-Compliant", "Compliance %"])
    for m in _build_members(db, coll.assessments):
        writer.writerow([m.group_code, m.workbook_name, m.product_type, m.items_total, m.items_compliant, m.items_non_compliant, f"{m.compliance_pct}%"])

    writer.writerow([])

    # Non-compliant items
    writer.writerow(["Non-Compliant Items"])
    writer.writerow(["Assessment", "Control Group", "Item", "Notes"])
    for item in get_non_compliant_items(db, coll.assessments):
        writer.writerow([item["assessment"], item["group_code"], item["item_text"], item["notes"]])

    return output.getvalue().encode("utf-8-sig")


# ── XLSX ───────────────────────────────────────────────────────────────────────

def export_xlsx(db: DBSession, coll: AssessmentCollection) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    stats = _compute_stats(coll.assessments)

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    BLUE = "4472C4"
    GREEN = "70AD47"
    RED = "C0392B"
    HEADER_FILL = PatternFill("solid", fgColor=BLUE)
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    def _hdr(cell, value):
        cell.value = value
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45

    _hdr(ws["A1"], "Collection")
    _hdr(ws["B1"], coll.name)
    ws["A2"].value = "Description"
    ws["B2"].value = coll.description or ""
    ws["A3"].value = "Product"
    ws["B3"].value = f"{coll.product_family} / {coll.product_type or ''}".strip(" /")
    ws["A4"].value = "Due Date"
    ws["B4"].value = str(coll.due_date) if coll.due_date else ""
    ws["A5"].value = "Status"
    ws["B5"].value = stats.status.replace("_", " ").title()
    ws["A6"].value = "Completion"
    ws["B6"].value = f"{stats.completion_pct}%"
    ws["A7"].value = "Compliance Score"
    ws["B7"].value = f"{stats.compliance_pct}%"
    ws["A8"].value = "Assessments"
    ws["B8"].value = f"{stats.started} / {stats.total} started"
    ws["A9"].value = "Non-Compliant Items"
    ws["B9"].value = stats.items_non_compliant

    # ── Assessments sheet ──────────────────────────────────────────────────────
    wa = wb.create_sheet("Assessments")
    headers = ["Control Group", "Workbook", "Product Type", "Items Total", "Compliant", "Non-Compliant", "Compliance %"]
    for col, h in enumerate(headers, 1):
        cell = wa.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row, m in enumerate(_build_members(db, coll.assessments), 2):
        wa.cell(row=row, column=1, value=m.group_code)
        wa.cell(row=row, column=2, value=m.workbook_name)
        wa.cell(row=row, column=3, value=m.product_type)
        wa.cell(row=row, column=4, value=m.items_total)
        wa.cell(row=row, column=5, value=m.items_compliant)
        wa.cell(row=row, column=6, value=m.items_non_compliant)
        pct_cell = wa.cell(row=row, column=7, value=f"{m.compliance_pct}%")
        if m.compliance_pct < 50:
            pct_cell.font = Font(color=RED, bold=True)
        elif m.compliance_pct >= 80:
            pct_cell.font = Font(color=GREEN, bold=True)
    for col in [1, 2, 3]:
        wa.column_dimensions[wa.cell(row=1, column=col).column_letter].width = 30
    for col in [4, 5, 6, 7]:
        wa.column_dimensions[wa.cell(row=1, column=col).column_letter].width = 16

    # ── Non-compliant items sheet ──────────────────────────────────────────────
    wn = wb.create_sheet("Non-Compliant Items")
    nc_headers = ["Assessment", "Control Group", "Item", "Notes"]
    for col, h in enumerate(nc_headers, 1):
        cell = wn.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="C0392B")
    for row, item in enumerate(get_non_compliant_items(db, coll.assessments), 2):
        wn.cell(row=row, column=1, value=item["assessment"])
        wn.cell(row=row, column=2, value=item["group_code"])
        wn.cell(row=row, column=3, value=item["item_text"])
        wn.cell(row=row, column=4, value=item["notes"])
    wn.column_dimensions["A"].width = 35
    wn.column_dimensions["B"].width = 18
    wn.column_dimensions["C"].width = 60
    wn.column_dimensions["D"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ────────────────────────────────────────────────────────────────────────

def export_pdf(db: DBSession, coll: AssessmentCollection) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    story = []

    BLUE = colors.HexColor("#4472C4")
    GREEN = colors.HexColor("#70AD47")
    RED = colors.HexColor("#C0392B")
    LIGHT_BLUE = colors.HexColor("#D6E4F7")

    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=16, textColor=BLUE, spaceAfter=4)
    h2_style = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12, textColor=BLUE, spaceAfter=4, spaceBefore=10)
    normal = styles["Normal"]

    stats = _compute_stats(coll.assessments)

    # Title
    story.append(Paragraph("Assessment Collection Report", title_style))
    story.append(Paragraph(coll.name, ParagraphStyle("sub", parent=styles["Normal"], fontSize=13, textColor=colors.grey, spaceAfter=8)))
    story.append(Spacer(1, 4*mm))

    # Summary table
    summary_data = [
        ["Product", f"{coll.product_family} / {coll.product_type or ''}".strip(" /")],
        ["Due Date", str(coll.due_date) if coll.due_date else "\u2014"],
        ["Status", stats.status.replace("_", " ").title()],
        ["Completion", f"{stats.completion_pct}% ({stats.started}/{stats.total} assessments)"],
        ["Compliance Score", f"{stats.compliance_pct}%"],
        ["Non-Compliant Items", str(stats.items_non_compliant)],
    ]
    summary_tbl = Table(summary_data, colWidths=[50*mm, 120*mm])
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), LIGHT_BLUE),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (1, 0), (1, -1), [colors.white]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(summary_tbl)
    story.append(Spacer(1, 6*mm))

    # Per-assessment scores
    story.append(Paragraph("Assessment Scores", h2_style))
    members = _build_members(db, coll.assessments)
    if members:
        tbl_data = [["Control Group", "Workbook", "Total", "Compliant", "Non-Compliant", "Score"]]
        for m in members:
            score_color = GREEN if m.compliance_pct >= 80 else (RED if m.compliance_pct < 50 else colors.orange)
            r = int(score_color.red * 255)
            g = int(score_color.green * 255)
            b = int(score_color.blue * 255)
            tbl_data.append([
                m.group_code,
                Paragraph(m.workbook_name, ParagraphStyle("small", parent=normal, fontSize=8)),
                str(m.items_total),
                str(m.items_compliant),
                str(m.items_non_compliant),
                Paragraph(f"<font color='#{r:02x}{g:02x}{b:02x}'><b>{m.compliance_pct}%</b></font>", ParagraphStyle("score", parent=normal, fontSize=9)),
            ])
        tbl = Table(tbl_data, colWidths=[22*mm, 65*mm, 16*mm, 16*mm, 22*mm, 18*mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(tbl)
    story.append(Spacer(1, 6*mm))

    # Non-compliant items
    nc_items = get_non_compliant_items(db, coll.assessments)
    if nc_items:
        story.append(Paragraph("Non-Compliant Items", h2_style))
        nc_data = [["Control", "Item", "Notes"]]
        for item in nc_items:
            nc_data.append([
                item["group_code"],
                Paragraph(item["item_text"][:200], ParagraphStyle("small", parent=normal, fontSize=8)),
                Paragraph(item["notes"][:150] if item["notes"] else "", ParagraphStyle("small", parent=normal, fontSize=8)),
            ])
        nc_tbl = Table(nc_data, colWidths=[22*mm, 100*mm, 47*mm])
        nc_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), RED),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF5F5")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(nc_tbl)

    # Footer
    story.append(Spacer(1, 10*mm))
    story.append(Paragraph(
        f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} \u00b7 ISMS CORE Platform",
        ParagraphStyle("footer", parent=normal, fontSize=7, textColor=colors.grey)
    ))

    doc.build(story)
    return buf.getvalue()
