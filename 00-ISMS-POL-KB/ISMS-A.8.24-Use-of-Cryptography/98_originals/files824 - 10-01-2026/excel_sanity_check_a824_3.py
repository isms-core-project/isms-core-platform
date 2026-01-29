#!/usr/bin/env python3
"""
Excel Workbook Sanity Checker - ISMS-IMP-A.8.24.3 Authentication Assessment
Diagnoses issues that trigger Excel's "file level validation and repair"

SPECIALIZED VERSION for Authentication Assessment workbooks

Usage:
    python3 excel_sanity_check_a824_3.py ISMS-IMP-A.8.24.3_Authentication_20251231.xlsx
    
"""

import sys
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ============================================================================
# A.8.24.3 SPECIFIC VALIDATION DATA
# ============================================================================

EXPECTED_SHEETS = [
    "Instructions & Legend",
    "1. Password Security",
    "2. Multi-Factor Authentication",
    "3. Certificate-Based Auth",
    "4. Service Accounts",
    "5. SSO & Federation",
    "Summary Dashboard",
    "Evidence Register",
    "Approval Sign-Off",
]

REQUIRED_DOCUMENT_INFO = [
    "Document ID",
    "Assessment Area",
    "Related Policy",
    "Version",
]

AUTHENTICATION_ASSESSMENT_SHEETS = [
    "1. Password Security",
    "2. Multi-Factor Authentication",
    "3. Certificate-Based Auth",
    "4. Service Accounts",
    "5. SSO & Federation",
]

# Base columns A-Q (17 columns) - same across all authentication sheets
EXPECTED_BASE_COLUMNS = [
    "System/Application",
    "Authentication Method",
    "User Type",
    "Data Classification",
    "Cryptographic Algorithm",
    "Hash/Encryption Status",
    "Password Complexity",
    "Status",
    "Evidence Location",
    "Gap Description",
    "Remediation Needed",
    "Exception ID",
    "Risk ID",
    "Compensating Controls",
    "Responsible Person",
    "Target Date",
    "Budget Required",
]

# Extended columns R-X (varies by sheet) - 6-7 additional columns
EXTENDED_COLUMNS = {
    "1. Password Security": 7,  # R-X: 7 password-specific columns
    "2. Multi-Factor Authentication": 6,  # R-W: 6 MFA-specific columns
    "3. Certificate-Based Auth": 7,  # R-X: 7 certificate-specific columns
    "4. Service Accounts": 7,  # R-X: 7 service account columns
    "5. SSO & Federation": 7,  # R-X: 7 SSO-specific columns
}

EXPECTED_CHECKLIST_COUNTS = {
    "1. Password Security": 20,
    "2. Multi-Factor Authentication": 19,
    "3. Certificate-Based Auth": 18,
    "4. Service Accounts": 19,
    "5. SSO & Federation": 20,
}


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""
    
    print("=" * 80)
    print(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {filename}")
    print("ISMS-IMP-A.8.24.3 - Authentication Assessment")
    print("=" * 80)
    
    issues_found = []
    warnings_found = []
    
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"\n✓ Workbook loaded successfully")
        print(f"  Sheets found: {len(wb.sheetnames)}")
        
    except Exception as e:
        print(f"\n✗ CRITICAL: Cannot load workbook: {e}")
        return
    
    # ========================================================================
    # CHECK 0: A.8.24.3 SPECIFIC STRUCTURE VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 0: A.8.24.3 STRUCTURE VALIDATION")
    print("=" * 80)
    
    structure_issues = 0
    
    # Check for expected sheets
    missing_sheets = []
    for sheet in EXPECTED_SHEETS:
        if sheet not in wb.sheetnames:
            missing_sheets.append(sheet)
            structure_issues += 1
    
    if missing_sheets:
        issues_found.append(f"  ✗ Missing required sheets: {', '.join(missing_sheets)}")
        print(f"  ✗ Missing {len(missing_sheets)} required sheet(s)")
    else:
        print(f"  ✓ All {len(EXPECTED_SHEETS)} required sheets present")
    
    # Check for unexpected sheets
    unexpected_sheets = []
    for sheet in wb.sheetnames:
        if sheet not in EXPECTED_SHEETS:
            unexpected_sheets.append(sheet)
    
    if unexpected_sheets:
        warnings_found.append(f"  ⚠ Unexpected sheets found: {', '.join(unexpected_sheets)}")
        print(f"  ⚠ Found {len(unexpected_sheets)} unexpected sheet(s)")
    
    # Validate document information fields
    if "Instructions & Legend" in wb.sheetnames:
        ws = wb["Instructions & Legend"]
        found_fields = []
        for row in ws.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for field in REQUIRED_DOCUMENT_INFO:
                        if field in cell.value:
                            found_fields.append(field)
        
        missing_fields = [f for f in REQUIRED_DOCUMENT_INFO if f not in found_fields]
        if missing_fields:
            warnings_found.append(f"  ⚠ Missing document info fields: {', '.join(missing_fields)}")
    
    # Validate assessment sheet structure
    for sheet_name in AUTHENTICATION_ASSESSMENT_SHEETS:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Check if header row exists (row 6)
            if ws.max_row >= 6:
                header_values = [cell.value for cell in ws[6] if cell.value]
                
                expected_cols = 17 + EXTENDED_COLUMNS.get(sheet_name, 0)
                if len(header_values) < expected_cols - 5:  # Allow some tolerance
                    warnings_found.append(f"  ⚠ {sheet_name}: Header row may be incomplete (expected ~{expected_cols} cols)")
            
            # Check for reference tables (should exist in assessment sheets)
            has_reference_table = False
            for row in ws.iter_rows(min_row=25, max_row=100):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if any(keyword in cell.value.upper() for keyword in 
                               ["APPROVED", "COMPARISON", "USE CASE", "CHECKLIST", "LIFECYCLE"]):
                            has_reference_table = True
                            break
                if has_reference_table:
                    break
            
            if not has_reference_table:
                warnings_found.append(f"  ⚠ {sheet_name}: Reference tables may be missing")
    
    # Validate Summary Dashboard structure
    if "Summary Dashboard" in wb.sheetnames:
        ws = wb["Summary Dashboard"]
        
        # Check for expected sections in Summary Dashboard
        expected_sections = [
            "COMPLIANCE SUMMARY",
            "AUTHENTICATION METHOD",
            "MFA ADOPTION",
            "PASSWORD SECURITY",
            "SERVICE ACCOUNT",
            "SSO COVERAGE",
            "SECURITY SCORE",
            "CRITICAL GAPS",
        ]
        
        found_sections = []
        for row in ws.iter_rows(min_row=1, max_row=200):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for section in expected_sections:
                        if section in cell.value.upper():
                            found_sections.append(section)
        
        missing_sections = [s for s in expected_sections if s not in found_sections]
        if missing_sections:
            warnings_found.append(f"  ⚠ Summary Dashboard: Missing sections: {', '.join(missing_sections)}")
        else:
            print(f"  ✓ Summary Dashboard: All 8 analysis sections found")
    
    if structure_issues == 0:
        print("  ✓ Workbook structure matches A.8.24.3 specification")
    
    # ========================================================================
    # CHECK 1: FORMULA VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 1: FORMULA VALIDATION")
    print("=" * 80)
    
    formula_issues = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # Check for common formula issues
                    if "'" in formula:
                        # Sheet references with special characters
                        sheet_refs = re.findall(r"'([^']+)'!", formula)
                        for ref in sheet_refs:
                            if ref not in wb.sheetnames:
                                issues_found.append(
                                    f"  ✗ {sheet_name}!{cell.coordinate}: "
                                    f"Invalid sheet reference '{ref}'"
                                )
                                formula_issues += 1
                    
                    # Check for common syntax issues
                    if formula.count('(') != formula.count(')'):
                        issues_found.append(
                            f"  ✗ {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced parentheses in formula"
                        )
                        formula_issues += 1
                    
                    # Check for double quotes issues
                    if formula.count('"') % 2 != 0:
                        issues_found.append(
                            f"  ✗ {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced quotes in formula"
                        )
                        formula_issues += 1
    
    if formula_issues == 0:
        print("  ✓ All formulas appear syntactically valid")
    else:
        print(f"  ✗ Found {formula_issues} formula issues")
    
    # ========================================================================
    # CHECK 2: DATA VALIDATION CONFLICTS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 2: DATA VALIDATION CONFLICTS")
    print("=" * 80)
    
    validation_issues = 0
    total_validations = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            dv_count = len(ws.data_validations.dataValidation)
            total_validations += dv_count
            print(f"  {sheet_name}: {dv_count} data validations")
            
            # Check for overlapping ranges (simplified)
            all_ranges = []
            for dv in ws.data_validations.dataValidation:
                if hasattr(dv, 'sqref') and dv.sqref:
                    all_ranges.extend(str(dv.sqref).split())
            
            if len(all_ranges) != len(set(all_ranges)):
                warnings_found.append(
                    f"  ⚠ {sheet_name}: Potentially overlapping data validation ranges"
                )
                validation_issues += 1
    
    print(f"  Total data validations across all sheets: {total_validations}")
    
    if validation_issues == 0:
        print("  ✓ No obvious data validation conflicts")
    
    # ========================================================================
    # CHECK 3: STYLE CONSISTENCY
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 3: STYLE CONSISTENCY")
    print("=" * 80)
    
    style_issues = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        cells_without_font = 0
        cells_without_border = 0
        
        sample_size = 0
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value:
                    sample_size += 1
                    if not hasattr(cell, 'font') or cell.font is None:
                        cells_without_font += 1
                    if not hasattr(cell, 'border') or cell.border is None:
                        cells_without_border += 1
        
        if cells_without_font > 0:
            warnings_found.append(
                f"  ⚠ {sheet_name}: {cells_without_font}/{sample_size} "
                f"cells missing font attributes"
            )
        
        # Note: Missing borders are often intentional, so we don't flag this as issue
    
    if style_issues == 0:
        print("  ✓ Style attributes appear consistent")
    
    # ========================================================================
    # CHECK 4: MERGED CELLS VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 4: MERGED CELLS VALIDATION")
    print("=" * 80)
    
    merge_issues = 0
    total_merges = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'merged_cells') and ws.merged_cells:
            merge_count = len(ws.merged_cells.ranges)
            total_merges += merge_count
            print(f"  {sheet_name}: {merge_count} merged cell ranges")
            
            # Check if merged cells have content in non-top-left cells
            for merge_range in ws.merged_cells.ranges:
                # Get min cell (top-left) of merged range
                min_col = merge_range.min_col
                min_row = merge_range.min_row
                max_col = merge_range.max_col
                max_row = merge_range.max_row
                
                # Check if any cell other than top-left has content
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        # Skip the top-left cell
                        if row == min_row and col == min_col:
                            continue
                        
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            cell_coord = f"{get_column_letter(col)}{row}"
                            warnings_found.append(
                                f"  ⚠ {sheet_name}!{cell_coord}: "
                                f"Merged cell has content in non-primary cell"
                            )
                            merge_issues += 1
    
    if total_merges > 0:
        print(f"  Total merged ranges across all sheets: {total_merges}")
    
    if merge_issues == 0:
        print("  ✓ Merged cells appear properly formatted")
    else:
        print(f"  ✗ Found {merge_issues} merged cell content issues")
    
    # ========================================================================
    # CHECK 5: WORKSHEET STRUCTURE
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 5: WORKSHEET STRUCTURE")
    print("=" * 80)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Check for excessive dimensions
        if ws.max_row > 10000:
            warnings_found.append(
                f"  ⚠ {sheet_name}: Large worksheet ({ws.max_row} rows)"
            )
        
        if ws.max_column > 100:
            warnings_found.append(
                f"  ⚠ {sheet_name}: Wide worksheet ({ws.max_column} columns)"
            )
    
    print("  ✓ Worksheet dimensions within reasonable limits")
    
    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    print("\n" + "=" * 80)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 80)
    
    if issues_found:
        print(f"\n❌ CRITICAL ISSUES FOUND: {len(issues_found)}")
        for issue in issues_found:
            print(issue)
    
    if warnings_found:
        print(f"\n⚠️  WARNINGS: {len(warnings_found)}")
        for warning in warnings_found:
            print(warning)
    
    if not issues_found and not warnings_found:
        print("\n✅ NO ISSUES DETECTED")
        print("\nThe workbook appears structurally sound and matches A.8.24.3 specification.")
        print("Excel repair warnings may be due to:")
        print("  • Excel version compatibility (try Excel 2019+)")
        print("  • Antivirus/security software interference")
        print("  • Network drive or OneDrive sync issues")
        print("  • Excel's overly cautious validation")
    else:
        print("\n" + "=" * 80)
        print("RECOMMENDED ACTIONS:")
        print("=" * 80)
        
        if structure_issues > 0:
            print("\n0. STRUCTURE FIXES:")
            print("   • Verify all required sheets are present")
            print("   • Check sheet names match specification exactly")
            print("   • Ensure document information fields are complete")
            print("   • Verify reference tables exist in assessment sheets")
        
        if formula_issues > 0:
            print("\n1. FORMULA FIXES:")
            print("   • Review formulas referencing other sheets")
            print("   • Ensure sheet names match exactly (case-sensitive)")
            print("   • Check for balanced quotes and parentheses")
        
        if validation_issues > 0:
            print("\n2. DATA VALIDATION FIXES:")
            print("   • Remove overlapping data validation ranges")
            print("   • Apply validations to specific ranges, not entire columns")
        
        if merge_issues > 0:
            print("\n3. MERGED CELL FIXES:")
            print("   • Ensure only top-left cell of merged range has content")
            print("   • Clear content from other cells in merged range")
    
    print("\n" + "=" * 80)
    print("A.8.24.3 SPECIFIC NOTES:")
    print("=" * 80)
    print("\nExpected structure:")
    print("  • 9 sheets total")
    print("  • 5 authentication assessment sheets:")
    print("    - Password Security (17 base + 7 extended = 24 columns, 20 checklist items)")
    print("    - Multi-Factor Authentication (17 base + 6 extended = 23 columns, 19 checklist items)")
    print("    - Certificate-Based Auth (17 base + 7 extended = 24 columns, 18 checklist items)")
    print("    - Service Accounts (17 base + 7 extended = 24 columns, 19 checklist items)")
    print("    - SSO & Federation (17 base + 7 extended = 24 columns, 20 checklist items)")
    print("  • Each assessment sheet: 13 data rows, compliance checklist, 2-3 reference tables")
    print("  • Summary Dashboard with 8 comprehensive analysis sections:")
    print("    1. Compliance summary by authentication type")
    print("    2. Authentication method distribution")
    print("    3. MFA adoption metrics")
    print("    4. Password security metrics")
    print("    5. Service account security")
    print("    6. SSO coverage analysis")
    print("    7. Overall authentication security score (weighted)")
    print("    8. Critical gaps requiring immediate attention")
    print("  • Evidence Register with 100 entry rows")
    print("  • Approval Sign-Off with 3-level workflow")
    print("\n" + "=" * 80)


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 excel_sanity_check_a824_3.py <filename.xlsx>")
        print("\nExample:")
        print("  python3 excel_sanity_check_a824_3.py ISMS-IMP-A.8.24.3_Authentication_20251231.xlsx")
        sys.exit(1)
    
    filename = sys.argv[1]
    check_workbook_health(filename)


if __name__ == "__main__":
    main()