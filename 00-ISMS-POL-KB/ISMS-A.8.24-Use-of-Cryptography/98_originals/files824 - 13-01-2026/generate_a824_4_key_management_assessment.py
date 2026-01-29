#!/usr/bin/env python3
"""
ISMS-IMP-A.8.24.4 - Key Management Assessment Excel Generator
ISO/IEC 27001:2022 Control A.8.24 (Use of Cryptography - Key Management)

Requirements:
    sudo apt install python3-openpyxl
    sudo apt install python3-pip

Usage:
    python3 generate_a824_4_key_management_assessment.py
    
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches markdown specification
    sheets = [
        "Instructions & Legend",
        "1. Key Generation",
        "2. Key Storage",
        "3. Key Rotation",
        "4. Key Backup & Recovery",
        "5. Certificate Management",
        "Summary Dashboard",
        "Evidence Register",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "exception_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
    }
    return styles


# ============================================================================
# SECTION 2: COLUMN DEFINITIONS (BASE + EXTENSIONS)
# ============================================================================

def get_base_key_mgmt_columns():
    """
    Return base 17 columns (A-Q) standard for ALL key management assessment sheets.
    
    Returns:
        dict: {column_name: width}
    """
    return {
        "System/Service": 30,
        "Key Type": 20,
        "Algorithm": 18,
        "Key Length (bits)": 16,
        "Storage Method": 22,
        "Data Classification": 18,
        "Lifecycle Stage": 18,
        "Status": 15,
        "Evidence Location": 28,
        "Gap Description": 30,
        "Remediation Needed": 14,
        "Exception ID": 14,
        "Risk ID": 14,
        "Compensating Controls": 30,
        "Responsible Person": 20,
        "Target Date": 14,
        "Budget Required": 15,
    }


def get_extended_columns(sheet_type):
    """
    Return additional columns (R-X) specific to each sheet type.
    
    Args:
        sheet_type: String identifier (generation, storage, rotation, backup, certificate)
    
    Returns:
        dict: {column_name: width} for columns R onwards
    """
    extensions = {
        "generation": {
            "Random Number Generator": 25,
            "Generation Location": 20,
            "Key Ceremony Performed": 16,
            "Entropy Source Verified": 16,
            "Generation Audit Log": 16,
        },
        "storage": {
            "HSM/KMS Model": 25,
            "FIPS 140 Level": 16,
            "Access Control Method": 22,
            "Encryption at Rest": 16,
            "Physical Security": 20,
            "Key Backup Exists": 16,
        },
        "rotation": {
            "Rotation Schedule": 20,
            "Last Rotation Date": 16,
            "Next Rotation Due": 16,
            "Automated Rotation": 16,
            "Rotation Tested": 16,
            "Re-encryption Required": 18,
        },
        "backup": {
            "Backup Method": 22,
            "Backup Encryption": 16,
            "Backup Location": 22,
            "Recovery Tested": 16,
            "Last Recovery Test": 16,
            "RTO (Recovery Time)": 18,
            "Escrow Agreement": 16,
        },
        "certificate": {
            "Certificate Type": 22,
            "Issuing CA": 25,
            "Certificate Validity": 16,
            "Expiration Date": 16,
            "Auto-Renewal": 16,
            "Revocation Method": 20,
            "Monitoring Enabled": 16,
        },
    }
    
    return extensions.get(sheet_type, {})


def get_all_columns(sheet_type):
    """Combine base + extended columns for a given sheet type."""
    base = get_base_key_mgmt_columns()
    extended = get_extended_columns(sheet_type)
    base.update(extended)
    return base


# ============================================================================
# SECTION 3: GLOBAL DATA VALIDATION DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create base dropdown validations common to all key management sheets.
    Returns dict of DataValidation objects.
    """
    validations = {}
    
    # Key Type
    validations['key_type'] = DataValidation(
        type="list",
        formula1='"Encryption Key,Signing Key,Authentication Key,Certificate,Master Key,Data Encryption Key,Key Encryption Key,Session Key,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['key_type'])
    
    # Algorithm
    validations['algorithm'] = DataValidation(
        type="list",
        formula1='"AES-256,AES-128,RSA-4096,RSA-3072,RSA-2048,Ed25519,ECDSA P-256,ECDSA P-384,ChaCha20,Other,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['algorithm'])
    
    # Key Length
    validations['key_length'] = DataValidation(
        type="list",
        formula1='"256,384,512,2048,3072,4096,8192,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['key_length'])
    
    # Storage Method
    validations['storage'] = DataValidation(
        type="list",
        formula1='"HSM,Cloud KMS,TPM,Software Keystore,File-based,Database,Smart Card,Vault,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['storage'])
    
    # Data Classification
    validations['data_class'] = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Restricted,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['data_class'])
    
    # Lifecycle Stage
    validations['lifecycle'] = DataValidation(
        type="list",
        formula1='"Active,Suspended,Archived,Destroyed,Pending Rotation,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['lifecycle'])
    
    # Status
    validations['status'] = DataValidation(
        type="list",
        formula1='"✅ Compliant,⚠️ Partial,❌ Non-Compliant,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['status'])
    
    # Remediation Needed
    validations['remediation'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(validations['remediation'])
    
    # Budget Required
    validations['budget'] = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=False
    )
    ws.add_data_validation(validations['budget'])
    
    # Response (Yes/No/Not Applicable)
    validations['response'] = DataValidation(
        type="list",
        formula1='"Yes,No,Not Applicable"',
        allow_blank=False
    )
    ws.add_data_validation(validations['response'])
    
    # Checklist Yes/No/N/A
    validations['checklist'] = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['checklist'])
    
    # Exception Yes/No
    validations['exception_yn'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(validations['exception_yn'])
    
    return validations


def create_extended_validations(ws, sheet_type):
    """Create sheet-specific dropdown validations for extended columns."""
    validations = {}
    
    if sheet_type == "generation":
        validations['rng'] = DataValidation(
            type="list",
            formula1='"Hardware RNG,OS Crypto RNG,HSM RNG,TPM RNG,FIPS 140-2 Validated,NIST SP 800-90A,Non-compliant,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['rng'])
        
        validations['gen_location'] = DataValidation(
            type="list",
            formula1='"HSM,KMS,TPM,Local System,Cloud Service,Smart Card,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['gen_location'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "storage":
        validations['fips_level'] = DataValidation(
            type="list",
            formula1='"Level 1,Level 2,Level 3,Level 4,Not Validated,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['fips_level'])
        
        validations['access_control'] = DataValidation(
            type="list",
            formula1='"Role-based,Certificate-based,MFA Required,PIN/Password,Smart Card,Dual Control,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['access_control'])
        
        validations['physical_sec'] = DataValidation(
            type="list",
            formula1='"Locked Data Center,Secure Room,Tamper-evident,Cloud Provider,Inadequate,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['physical_sec'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "rotation":
        validations['rotation_sched'] = DataValidation(
            type="list",
            formula1='"Hourly,Daily,Weekly,Monthly,Quarterly,Annually,Biennially,On Compromise,Never,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['rotation_sched'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
        
        validations['reencrypt'] = DataValidation(
            type="list",
            formula1='"Yes,No,In Progress,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['reencrypt'])
    
    elif sheet_type == "backup":
        validations['backup_method'] = DataValidation(
            type="list",
            formula1='"HSM Backup,KMS Backup,Encrypted File,Key Escrow,Split Knowledge,Cloud Backup,No Backup,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['backup_method'])
        
        validations['recovery_tested'] = DataValidation(
            type="list",
            formula1='"Yes,No,Scheduled,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['recovery_tested'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    elif sheet_type == "certificate":
        validations['cert_type'] = DataValidation(
            type="list",
            formula1='"TLS/SSL Server,TLS/SSL Client,Code Signing,Email (S/MIME),User Authentication,CA Certificate,Wildcard,SAN,Self-signed,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['cert_type'])
        
        validations['cert_validity'] = DataValidation(
            type="list",
            formula1='"≤90 days,91-180 days,181-365 days,366-397 days,>397 days (non-compliant),Expired,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['cert_validity'])
        
        validations['revocation'] = DataValidation(
            type="list",
            formula1='"CRL,OCSP,OCSP Stapling,Not Configured,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['revocation'])
        
        validations['yn_na'] = DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(validations['yn_na'])
    
    return validations


# ============================================================================
# END OF PART 1
# ============================================================================

# ============================================================================
# SECTION 4: CHECKLIST DEFINITIONS PER KEY MANAGEMENT AREA
# ============================================================================

def get_checklist_items(sheet_type):
    """
    Return checklist items for each key management assessment type.
    
    Args:
        sheet_type: String identifier (generation, storage, rotation, backup, certificate)
    
    Returns:
        list: Checklist items as strings
    """
    checklists = {
        "generation": [
            "Approved algorithms only (AES-256, RSA ≥3072, Ed25519, ECDSA P-256+)",
            "Cryptographically secure random number generator (CSRNG) used",
            "FIPS 140-2 Level 2+ validated cryptographic module (production)",
            "Key generation performed in secure environment (HSM/KMS/TPM)",
            "Weak algorithms prohibited (DES, 3DES, MD5, SHA-1, RSA <2048)",
            "Key ceremony documented for master/root keys",
            "Dual control/split knowledge for critical key generation",
            "Key generation audit logging enabled",
            "Generated keys never exposed in plaintext outside secure boundary",
            "Entropy source quality verified and tested",
            "Key generation procedures documented",
            "Key metadata tracked (creation date, purpose, owner, classification)",
            "Test/development keys distinct from production keys",
            "Default/vendor keys replaced before production use",
        ],
        "storage": [
            "Production keys stored in HSM, cloud KMS, or TPM",
            "FIPS 140-2 Level 2+ (or FIPS 140-3 equivalent) for production",
            "Keys encrypted when stored (key encryption keys protected by HSM)",
            "Keys never stored in plaintext files or code repositories",
            "Keys never stored in application configuration files",
            "Keys never stored in environment variables (production)",
            "Keys never stored in databases without encryption",
            "Access to keys restricted by role-based access control (RBAC)",
            "Multi-factor authentication required for key access",
            "Dual control/split knowledge for master keys",
            "Key access logged and monitored",
            "Keys physically secured (locked facilities, tamper-evident HSMs)",
            "Key storage location geographically documented",
            "Keys stored separately from encrypted data",
            "Backup keys stored with equal or greater security",
            "Key export restricted and logged",
            "Unused/obsolete keys securely destroyed",
        ],
        "rotation": [
            "Key rotation policy documented and approved",
            "Encryption keys (data at rest) rotated annually minimum",
            "Signing keys rotated biennially minimum",
            "Session keys rotated hourly/daily",
            "Master keys rotated every 1-3 years",
            "Key rotation automated where possible",
            "Key rotation triggers defined (time, usage, compromise)",
            "Data re-encryption performed after key rotation",
            "Old keys retained for decryption (with restricted access)",
            "Key version tracking implemented",
            "Rotation procedures tested in non-production",
            "Rotation audit logs maintained",
            "Emergency rotation procedure documented",
            "Rotation scheduling prevents service disruption",
            "Key rotation monitored and alerts configured",
            "Cryptoperiod limits enforced (max key usage duration)",
        ],
        "backup": [
            "Key backup and recovery procedures documented",
            "All critical keys have backup copies",
            "Backup keys encrypted with equal or stronger encryption",
            "Backup keys stored separately from production keys",
            "Geographic separation of backup keys (offsite/different region)",
            "Backup key access restricted (dual control/escrow)",
            "Key recovery procedures tested quarterly",
            "Recovery time objective (RTO) defined and documented",
            "Recovery point objective (RPO) defined and documented",
            "Key escrow agreements in place (where required)",
            "Split knowledge/dual control for backup key recovery",
            "Backup key inventory maintained and audited",
            "Backup keys versioned and synchronized with production",
            "Emergency recovery procedures documented",
            "Key recovery audit logging enabled",
            "Backup keys updated when production keys rotate",
            "Cloud KMS backup to different region/account",
            "HSM backup cards stored in secure facility",
        ],
        "certificate": [
            "Certificates issued from trusted public CA (for external services)",
            "Internal CA infrastructure properly secured (for internal services)",
            "Certificate validity ≤397 days (public TLS certificates)",
            "Certificate validity ≤90 days preferred (automated renewal)",
            "Self-signed certificates NOT used in production",
            "Certificate inventory maintained and up to date",
            "Certificate expiration monitoring implemented",
            "Alerts configured for 30/60/90 days before expiration",
            "Automated certificate renewal configured",
            "Certificate renewal tested in non-production",
            "Certificate revocation procedures documented",
            "CRL or OCSP configured for revocation checking",
            "Compromised certificate revocation within 24 hours",
            "Certificate private keys stored securely (HSM/KMS preferred)",
            "Certificate private keys never shared across systems",
            "Wildcard certificates use documented and restricted",
            "Certificate signing requests (CSR) process documented",
            "Certificate audit logs maintained",
            "Code signing certificates protected with hardware tokens",
            "Email certificates (S/MIME) issued per user",
        ],
    }
    
    return checklists.get(sheet_type, [])


# ============================================================================
# SECTION 5: REFERENCE TABLE DEFINITIONS PER SHEET TYPE
# ============================================================================

def get_reference_tables(sheet_type):
    """
    Return reference table definitions for each sheet type.
    
    Returns:
        list: [(table_title, headers, data_rows), ...]
    """
    tables = {
        "generation": [
            (
                "APPROVED ALGORITHMS BY USE CASE",
                ["Use Case", "Minimum Algorithm", "Key Length", "Notes"],
                [
                    ("Symmetric Encryption (Data at Rest)", "AES", "256 bits", "AES-256-GCM preferred"),
                    ("Symmetric Encryption (Data in Transit)", "AES, ChaCha20", "256 bits", "ChaCha20-Poly1305 for mobile"),
                    ("Asymmetric Encryption", "RSA, ECDH", "RSA ≥3072, ECDH P-256+", "RSA-4096 for long-term"),
                    ("Digital Signatures", "RSA, Ed25519, ECDSA", "RSA ≥3072, Ed25519, ECDSA P-256+", "Ed25519 preferred for new systems"),
                    ("Key Exchange", "ECDH, RSA", "ECDH P-256+, RSA ≥3072", "Use ephemeral keys for PFS"),
                    ("Hashing", "SHA-256, SHA-384, SHA-512", "N/A (output size)", "SHA-256 minimum"),
                    ("HMAC", "HMAC-SHA256", "≥256 bits", "For message authentication"),
                ]
            ),
        ],
        "storage": [
            (
                "APPROVED KEY STORAGE SOLUTIONS",
                ["Solution Type", "Examples", "FIPS 140-2 Level", "Use Case"],
                [
                    ("Hardware HSM", "Thales Luna, Entrust nShield, Utimaco", "Level 3/4", "High-security production"),
                    ("Cloud HSM", "AWS CloudHSM, Azure Dedicated HSM", "Level 3", "Cloud-native high-security"),
                    ("Cloud KMS", "AWS KMS, Azure Key Vault, GCP Cloud KMS", "Level 2/3", "General cloud workloads"),
                    ("TPM", "TPM 2.0 (various vendors)", "Level 2", "Endpoint/device keys"),
                    ("Software Vault", "HashiCorp Vault, CyberArk", "Varies", "Non-production acceptable"),
                ]
            ),
        ],
        "rotation": [
            (
                "ROTATION SCHEDULE BY KEY TYPE",
                ["Key Type", "Minimum Rotation Frequency", "Recommended Frequency", "Trigger Events"],
                [
                    ("Master Encryption Key", "1 year", "1 year", "Compromise, personnel change"),
                    ("Data Encryption Key", "1 year", "90 days", "High volume usage, compromise"),
                    ("Key Encryption Key", "1 year", "1 year", "Master key rotation, compromise"),
                    ("Signing Key", "2 years", "1 year", "Certificate expiry, compromise"),
                    ("Session Key", "24 hours", "1 hour", "Session end, compromise"),
                    ("API Key", "90 days", "30-90 days", "Personnel change, compromise"),
                    ("TLS/SSL Certificate Private Key", "Certificate validity", "Annual", "Certificate renewal, compromise"),
                    ("SSH Key", "1 year", "90 days", "Personnel change, compromise"),
                    ("Database TDE Key", "1 year", "1 year", "Compliance requirement, compromise"),
                ]
            ),
        ],
        "backup": [
            (
                "BACKUP STRATEGIES BY KEY TYPE",
                ["Key Type", "Recommended Backup Method", "Storage Location", "Recovery Testing Frequency"],
                [
                    ("Master Keys", "HSM backup card + escrow", "Offsite secure vault", "Annually"),
                    ("Encryption Keys", "KMS automated backup", "Different region/account", "Quarterly"),
                    ("Certificate Private Keys", "Encrypted file backup", "Secure file storage", "Semi-annually"),
                    ("SSH Keys", "Encrypted backup + key management system", "Secrets vault", "Annually"),
                    ("API Keys", "Secrets management system", "Multi-region vault", "Quarterly"),
                    ("Database TDE Keys", "Native database backup + KMS", "Encrypted backup storage", "Quarterly"),
                ]
            ),
        ],
        "certificate": [
            (
                "CERTIFICATE VALIDITY STANDARDS",
                ["Certificate Type", "Maximum Validity", "Recommended Validity", "Renewal Process"],
                [
                    ("Public TLS/SSL", "397 days (13 months)", "90 days", "ACME/Let's Encrypt automated"),
                    ("Internal TLS/SSL", "2 years", "1 year", "Internal CA automated"),
                    ("Code Signing", "3 years", "1 year", "Manual with hardware token"),
                    ("Email (S/MIME)", "2 years", "1 year", "PKI automated or manual"),
                    ("Client Authentication", "2 years", "1 year", "PKI automated"),
                    ("Root CA", "20 years", "10 years", "Manual with ceremony"),
                    ("Intermediate CA", "10 years", "5 years", "Manual with ceremony"),
                ]
            ),
            (
                "CERTIFICATE LIFECYCLE STAGES",
                ["Stage", "Actions", "Responsible Role", "Documentation Required"],
                [
                    ("Request", "Generate CSR, validate domain/identity", "System Owner", "CSR details, approval"),
                    ("Issuance", "CA validation, certificate generation", "CA Admin", "Issuance record, validation log"),
                    ("Installation", "Deploy certificate to service", "System Admin", "Deployment log, verification"),
                    ("Monitoring", "Track expiration, validate configuration", "Security Team", "Monitoring dashboard"),
                    ("Renewal", "Generate new CSR, request renewal", "System Owner", "Renewal approval, testing"),
                    ("Revocation", "Revoke compromised/obsolete certificate", "CA Admin", "Incident record, revocation log"),
                    ("Archival", "Archive expired certificates", "Security Team", "Archive log, retention period"),
                ]
            ),
        ],
    }
    
    return tables.get(sheet_type, [])


# ============================================================================
# END OF PART 2
# ============================================================================

# ============================================================================
# SECTION 6: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create the Instructions & Legend sheet matching markdown spec."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.8.24.4 – Key Management Assessment\n"
        "ISO/IEC 27001:2022 - Control A.8.24: Use of Cryptography"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.8.24.4"),
        ("Assessment Area", "Key Management Cryptographic Controls"),
        ("Related Policy", "ISMS-POL-A.8.24-S2.4"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organization", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Complete each worksheet tab (1–5) for applicable key management areas",
        "2. Use dropdown menus for standardized entries (Status, Key Type, Algorithm, Storage Method, etc.)",
        "3. Fill in yellow-highlighted cells with your information",
        "4. If Status = Partial or Non-Compliant, complete the Exception/Deviation section",
        "5. Document key lifecycle details for each cryptographic system",
        "6. Provide evidence location/path for each implementation entry",
        "7. Summary Dashboard auto-calculates compliance statistics per key management area",
        "8. Maintain the Evidence Register for audit traceability",
        "9. Obtain final approval and sign-off in the Approval Sign-Off sheet",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws[f"A{row+1}"] = "STATUS LEGEND"
    ws[f"A{row+1}"].font = Font(bold=True, size=12)

    legend = [
        ("Symbol", "Status", "Description"),
        ("✅", "Compliant", "Fully meets policy requirements"),
        ("⚠️", "Partial", "Some requirements met, gaps exist"),
        ("❌", "Non-Compliant", "Does not meet policy requirements"),
        ("N/A", "Not Applicable", "Requirement does not apply"),
    ]

    row += 2
    header_row = row
    for c, h in enumerate(legend[0], start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for sym, status, desc in legend[1:]:
        ws.cell(row=row, column=1, value=sym).border = styles["border"]
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if status == "Compliant":
            s.fill = styles["status_compliant"]["fill"]
        elif status == "Partial":
            s.fill = styles["status_partial"]["fill"]
        elif status == "Non-Compliant":
            s.fill = styles["status_noncompliant"]["fill"]

        row += 1

    ws[f"A{row+2}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row+2}"].font = Font(bold=True, size=12)

    evidence_types = [
        "✓ Key generation audit logs",
        "✓ HSM configuration and initialization records",
        "✓ KMS architecture diagrams",
        "✓ Key rotation schedules and logs",
        "✓ Key backup and recovery procedures",
        "✓ Key escrow agreements",
        "✓ Certificate lifecycle management reports",
        "✓ Certificate Authority (CA) audit reports",
        "✓ Key ceremony documentation",
        "✓ Cryptographic module validation certificates (FIPS 140-2/3)",
        "✓ Key usage and access logs",
        "✓ Key destruction records",
        "✓ Algorithm compliance verification reports",
        "✓ Third-party key management audit reports",
    ]
    row += 3
    for e in evidence_types:
        ws[f"A{row}"] = e
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: GENERIC ASSESSMENT SHEET ENGINE
# ============================================================================

def create_assessment_sheet(ws, styles, section_title, policy_ref, question, 
                           sheet_type):
    """
    Generic assessment sheet creator for key management controls.
    
    Args:
        ws: worksheet object
        styles: style dictionary
        section_title: e.g., "1. KEY GENERATION - CRYPTOGRAPHIC KEY CREATION"
        policy_ref: policy requirement text
        question: assessment question
        sheet_type: key for columns/checklist/tables (generation, storage, rotation, backup, certificate)
    """
    columns = get_all_columns(sheet_type)
    checklist_items = get_checklist_items(sheet_type)
    reference_tables = get_reference_tables(sheet_type)
    
    total_cols = len(columns)
    last_col_letter = get_column_letter(total_cols)
    
    # ---------- HEADER ----------
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"{section_title}\nPolicy Requirement: {policy_ref}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # ---------- ASSESSMENT QUESTION ----------
    ws.merge_cells(f"A3:{last_col_letter}3")
    ws["A3"] = question
    ws["A3"].font = Font(bold=True, size=11)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # ---------- RESPONSE DROPDOWN ----------
    ws["A4"] = "Response:"
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    # Create validations
    base_validations = create_base_validations(ws)
    extended_validations = create_extended_validations(ws, sheet_type)
    base_validations['response'].add(ws["B4"])

    # ---------- COLUMN HEADERS ----------
    header_row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ---------- EXAMPLE ROW ----------
    example_row = 7
    example_values = _get_example_row(sheet_type, total_cols)
    for col_idx, value in enumerate(example_values, start=1):
        cell = ws.cell(row=example_row, column=col_idx, value=value)
        cell.font = Font(italic=True, color="808080")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---------- DATA ENTRY ROWS (8-30) ----------
    start_row = 8
    end_row = 30
    
    for r in range(start_row, end_row + 1):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # Apply base validations to standard columns (B-Q, indices 2-17)
    for r in range(start_row, end_row + 1):
        base_validations['key_type'].add(ws.cell(row=r, column=2))      # B: Key Type
        base_validations['algorithm'].add(ws.cell(row=r, column=3))     # C: Algorithm
        base_validations['key_length'].add(ws.cell(row=r, column=4))    # D: Key Length
        base_validations['storage'].add(ws.cell(row=r, column=5))       # E: Storage Method
        base_validations['data_class'].add(ws.cell(row=r, column=6))    # F: Data Classification
        base_validations['lifecycle'].add(ws.cell(row=r, column=7))     # G: Lifecycle Stage
        base_validations['status'].add(ws.cell(row=r, column=8))        # H: Status
        base_validations['remediation'].add(ws.cell(row=r, column=11))  # K: Remediation Needed
        base_validations['budget'].add(ws.cell(row=r, column=17))       # Q: Budget Required

    # Apply extended validations for columns R onwards (18+)
    _apply_extended_validations(ws, sheet_type, extended_validations, start_row, end_row)

    ws.freeze_panes = "A7"

    next_row = end_row + 2

    # ---------- COMPLIANCE CHECKLIST ----------
    ws[f"A{next_row}"] = f"{section_title.split('-')[0].strip().upper()} CHECKLIST"
    ws[f"A{next_row}"].font = Font(bold=True, size=11)
    next_row += 1

    ws[f"A{next_row}"] = "☐"
    ws[f"B{next_row}"] = "Requirement"
    ws[f"C{next_row}"] = "Status"
    for col in ["A", "B", "C"]:
        ws[f"{col}{next_row}"].font = Font(bold=True)
    next_row += 1

    checklist_start = next_row
    for item in checklist_items:
        ws[f"A{next_row}"] = "☐"
        ws[f"B{next_row}"] = item
        ws[f"C{next_row}"].fill = styles["input_cell"]["fill"]
        ws[f"C{next_row}"].border = styles["border"]

        base_validations['checklist'].add(ws[f"C{next_row}"])
        next_row += 1

    # Checklist score
    ws[f"A{next_row}"] = "Checklist Score:"
    ws[f"A{next_row}"].font = Font(bold=True)
    ws[f"B{next_row}"] = (
        f'=COUNTIF(C{checklist_start}:C{next_row-1},"Yes")/'
        f'COUNTA(C{checklist_start}:C{next_row-1})*100&"%"'
    )
    ws[f"B{next_row}"].font = Font(bold=True, color="0000FF")
    next_row += 2

    # ---------- REFERENCE TABLES ----------
    for table_title, headers, data_rows in reference_tables:
        ws[f"A{next_row}"] = table_title
        ws[f"A{next_row}"].font = Font(bold=True, size=11)
        next_row += 1
        
        # Table headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=header)
            cell.font = styles["column_header"]["font"]
            cell.fill = styles["column_header"]["fill"]
            cell.alignment = styles["column_header"]["alignment"]
            cell.border = styles["column_header"]["border"]
        next_row += 1
        
        # Table data rows
        for data_row in data_rows:
            for col_idx, value in enumerate(data_row, start=1):
                cell = ws.cell(row=next_row, column=col_idx, value=value)
                cell.border = styles["border"]
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            next_row += 1
        
        next_row += 1  # Space between tables

    # ---------- EXCEPTION/DEVIATION BLOCK ----------
    ws.merge_cells(f"A{next_row}:{last_col_letter}{next_row}")
    ws[f"A{next_row}"] = "EXCEPTION / DEVIATION DOCUMENTATION (Complete if Status = Partial or Non-Compliant)"
    ws[f"A{next_row}"].font = styles["exception_header"]["font"]
    ws[f"A{next_row}"].fill = styles["exception_header"]["fill"]
    ws[f"A{next_row}"].alignment = styles["exception_header"]["alignment"]

    next_row += 1
    exception_fields = [
        ("Formal exception request submitted:", "Yes,No"),
        ("Exception ID:", ""),
        ("Risk acceptance documented:", "Yes,No"),
        ("Risk ID:", ""),
        ("Compensating Controls (summary):", ""),
        ("☐ Enhanced monitoring and alerting", ""),
        ("☐ Restricted network access / air-gapped systems", ""),
        ("☐ Additional encryption layer", ""),
        ("☐ Manual key management procedures", ""),
        ("☐ Increased audit frequency", ""),
        ("☐ Physical security controls", ""),
        ("☐ Other (describe):", ""),
    ]

    for label, options in exception_fields:
        ws[f"A{next_row}"] = label
        ws[f"A{next_row}"].font = Font(bold=True)
        ws.merge_cells(f"B{next_row}:D{next_row}")
        ws[f"B{next_row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{next_row}"].border = styles["border"]
        ws[f"B{next_row}"].alignment = styles["input_cell"]["alignment"]

        if options:
            base_validations['exception_yn'].add(ws[f"B{next_row}"])

        next_row += 1

    # ---------- NOTES ----------
    next_row += 1
    ws.merge_cells(f"A{next_row}:{last_col_letter}{next_row}")
    ws[f"A{next_row}"] = "ADDITIONAL NOTES / COMMENTS"
    ws[f"A{next_row}"].font = Font(bold=True, size=11)

    next_row += 1
    ws.merge_cells(f"A{next_row}:{last_col_letter}{next_row+8}")
    ws[f"A{next_row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{next_row}"].border = styles["border"]
    ws[f"A{next_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Column widths for checklist/tables
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 20

    return (start_row, end_row, 8)  # Return status column index (H = 8)


def _get_example_row(sheet_type, total_cols):
    """Generate example row values based on sheet type."""
    examples = {
        "generation": [
            "Example: Production Database Encryption - FINDB01", "Data Encryption Key", "AES-256", "256",
            "HSM", "Restricted", "Active", "✅ Compliant",
            "/evidence/key-mgmt/FINDB01-key-generation-log.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "HSM RNG", "HSM", "Yes", "Yes", "Yes"
        ],
        "storage": [
            "Example: Azure Key Vault - App1 Keys", "Key Encryption Key", "RSA-4096", "4096",
            "Cloud KMS", "Confidential", "Active", "✅ Compliant",
            "/evidence/key-mgmt/app1-kv-config.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "Azure Key Vault", "Level 2", "MFA Required", "Yes", "Cloud Provider", "Yes"
        ],
        "rotation": [
            "Example: API Signing Keys", "Signing Key", "RSA-3072", "3072",
            "Cloud KMS", "Confidential", "Active", "✅ Compliant",
            "/evidence/key-mgmt/api-key-rotation.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "Quarterly", "2024-10-01", "2025-01-01", "Yes", "Yes", "No"
        ],
        "backup": [
            "Example: Master Encryption Key Backup", "Master Key", "AES-256", "256",
            "HSM", "Restricted", "Active", "✅ Compliant",
            "/evidence/key-mgmt/master-key-backup.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "HSM Backup", "Yes", "Offsite Secure Vault", "Yes", "2024-06-15", "<1 hour", "Yes"
        ],
        "certificate": [
            "Example: Production Web Server TLS", "Certificate", "RSA-2048", "2048",
            "TPM", "Public", "Active", "✅ Compliant",
            "/evidence/key-mgmt/web-tls-cert.pdf", "None", "No", "N/A", "N/A", "N/A", "N/A", "N/A", "No",
            "TLS/SSL Server", "Let's Encrypt", "≤90 days", "2025-03-15", "Yes", "OCSP", "Yes"
        ],
    }
    
    base_example = examples.get(sheet_type, ["Example"] * total_cols)
    # Pad or trim to match total_cols
    if len(base_example) < total_cols:
        base_example.extend(["N/A"] * (total_cols - len(base_example)))
    return base_example[:total_cols]


def _apply_extended_validations(ws, sheet_type, validations, start_row, end_row):
    """Apply extended column validations based on sheet type."""
    if sheet_type == "generation":
        for r in range(start_row, end_row + 1):
            validations['rng'].add(ws.cell(row=r, column=18))        # R
            validations['gen_location'].add(ws.cell(row=r, column=19)) # S
            validations['yn_na'].add(ws.cell(row=r, column=20))      # T: Key Ceremony
            validations['yn_na'].add(ws.cell(row=r, column=21))      # U: Entropy Verified
            validations['yn_na'].add(ws.cell(row=r, column=22))      # V: Audit Log
    
    elif sheet_type == "storage":
        for r in range(start_row, end_row + 1):
            # R: HSM/KMS Model - free text
            validations['fips_level'].add(ws.cell(row=r, column=19))    # S
            validations['access_control'].add(ws.cell(row=r, column=20)) # T
            validations['yn_na'].add(ws.cell(row=r, column=21))         # U: Encryption at Rest
            validations['physical_sec'].add(ws.cell(row=r, column=22))  # V
            validations['yn_na'].add(ws.cell(row=r, column=23))         # W: Key Backup Exists
    
    elif sheet_type == "rotation":
        for r in range(start_row, end_row + 1):
            validations['rotation_sched'].add(ws.cell(row=r, column=18)) # R
            # S: Last Rotation Date - date picker
            # T: Next Rotation Due - date picker
            validations['yn_na'].add(ws.cell(row=r, column=21))         # U: Automated Rotation
            validations['yn_na'].add(ws.cell(row=r, column=22))         # V: Rotation Tested
            validations['reencrypt'].add(ws.cell(row=r, column=23))     # W
    
    elif sheet_type == "backup":
        for r in range(start_row, end_row + 1):
            validations['backup_method'].add(ws.cell(row=r, column=18)) # R
            validations['yn_na'].add(ws.cell(row=r, column=19))         # S: Backup Encryption
            # T: Backup Location - free text
            validations['recovery_tested'].add(ws.cell(row=r, column=21)) # U
            # V: Last Recovery Test - date picker
            # W: RTO - free text
            validations['yn_na'].add(ws.cell(row=r, column=24))         # X: Escrow Agreement
    
    elif sheet_type == "certificate":
        for r in range(start_row, end_row + 1):
            validations['cert_type'].add(ws.cell(row=r, column=18))     # R
            # S: Issuing CA - free text
            validations['cert_validity'].add(ws.cell(row=r, column=20)) # T
            # U: Expiration Date - date picker
            validations['yn_na'].add(ws.cell(row=r, column=22))         # V: Auto-Renewal
            validations['revocation'].add(ws.cell(row=r, column=23))    # W
            validations['yn_na'].add(ws.cell(row=r, column=24))         # X: Monitoring Enabled


# ============================================================================
# END OF PART 3
# ============================================================================

# ============================================================================
# SECTION 8: INDIVIDUAL ASSESSMENT SHEET DEFINITIONS
# ============================================================================

def create_1_key_generation(ws, styles):
    """1. Key Generation - matches markdown spec exactly."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="1. KEY GENERATION - CRYPTOGRAPHIC KEY CREATION",
        policy_ref="Cryptographic keys MUST be generated using approved algorithms with cryptographically secure random number generators (Policy Section 2.4.1)",
        question="Does your organization generate cryptographic keys for encryption, signing, or authentication purposes?",
        sheet_type="generation",
    )


def create_2_key_storage(ws, styles):
    """2. Key Storage - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="2. KEY STORAGE - SECURE KEY PROTECTION",
        policy_ref="Cryptographic keys MUST be stored in hardware security modules (HSM), cloud KMS, or TPM for production systems. Software keystores permitted only for non-production with documented risk acceptance (Policy Section 2.4.2)",
        question="Does your organization store cryptographic keys that require protection?",
        sheet_type="storage",
    )


def create_3_key_rotation(ws, styles):
    """3. Key Rotation - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="3. KEY ROTATION - CRYPTOGRAPHIC KEY LIFECYCLE MANAGEMENT",
        policy_ref="Cryptographic keys MUST be rotated according to defined schedules: Encryption keys annually, signing keys biennially, session keys hourly/daily (Policy Section 2.4.3)",
        question="Does your organization have a key rotation policy and implementation for cryptographic keys?",
        sheet_type="rotation",
    )


def create_4_key_backup_recovery(ws, styles):
    """4. Key Backup & Recovery - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="4. KEY BACKUP & RECOVERY - BUSINESS CONTINUITY",
        policy_ref="Cryptographic keys MUST have documented backup and recovery procedures. Backup keys stored with equal or greater security than production keys (Policy Section 2.4.4)",
        question="Does your organization maintain backup and recovery procedures for cryptographic keys?",
        sheet_type="backup",
    )


def create_5_certificate_management(ws, styles):
    """5. Certificate Management - matches markdown spec."""
    create_assessment_sheet(
        ws=ws,
        styles=styles,
        section_title="5. CERTIFICATE MANAGEMENT - DIGITAL CERTIFICATE LIFECYCLE",
        policy_ref="Digital certificates MUST be issued from trusted Certificate Authorities. Certificate lifecycle management includes issuance, renewal, revocation, and expiration monitoring (Policy Section 2.4.5)",
        question="Does your organization use digital certificates for TLS/SSL, code signing, email signing, or authentication?",
        sheet_type="certificate",
    )


# ============================================================================
# SECTION 9: SUMMARY DASHBOARD (11 ANALYSIS SECTIONS)
# ============================================================================

def create_summary_dashboard(ws, styles):
    """Create comprehensive summary dashboard with 11 analysis sections."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "KEY MANAGEMENT ASSESSMENT - COMPLIANCE SUMMARY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # ---------- 1. COMPLIANCE SUMMARY TABLE ----------
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    areas = [
        ("1. Key Generation", "1. Key Generation"),
        ("2. Key Storage", "2. Key Storage"),
        ("3. Key Rotation", "3. Key Rotation"),
        ("4. Key Backup & Recovery", "4. Key Backup & Recovery"),
        ("5. Certificate Management", "5. Certificate Management"),
    ]

    row += 1
    start_data_row = row
    for label, sheet in areas:
        ws.cell(row=row, column=1, value=label)
        
        status_range = f"'{sheet}'!H8:H30"

        ws.cell(row=row, column=2, value=f'=COUNTA({status_range})')
        ws.cell(row=row, column=3, value=f'=COUNTIF({status_range},"✅ Compliant")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({status_range},"⚠️ Partial")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({status_range},"❌ Non-Compliant")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({status_range},"N/A")')
        ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    # Total row
    total_row = row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{row-1})")
        cell.font = Font(bold=True)

    total_pct = ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
    total_pct.font = Font(bold=True, color="0000FF", size=12)

    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 15

    # ---------- 2. KEY MANAGEMENT MATURITY BY AREA ----------
    row += 3
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "KEY MANAGEMENT MATURITY BY AREA"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    maturity_headers = ["Key Management Area", "Checklist Score", "Maturity Level"]
    for col_idx, header in enumerate(maturity_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    maturity_areas = [
        ("Key Generation", "1. Key Generation"),
        ("Key Storage", "2. Key Storage"),
        ("Key Rotation", "3. Key Rotation"),
        ("Key Backup & Recovery", "4. Key Backup & Recovery"),
        ("Certificate Management", "5. Certificate Management"),
    ]

    row += 1
    for area, sheet in maturity_areas:
        ws.cell(row=row, column=1, value=area)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual/formula reference to checklist
        ws.cell(row=row, column=3, value=f'=IF(B{row}>=90%,"Advanced",IF(B{row}>=70%,"Intermediate","Basic"))')
        row += 1

    # ---------- 3. ALGORITHM DISTRIBUTION ANALYSIS ----------
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "ALGORITHM DISTRIBUTION ANALYSIS"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    algo_headers = ["Algorithm Type", "Count", "Percentage", "Compliance Status"]
    for col_idx, header in enumerate(algo_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    algorithms = [
        ("AES-256", "Approved"),
        ("AES-128", "Approved"),
        ("RSA-4096", "Approved"),
        ("RSA-3072", "Approved"),
        ("RSA-2048", "Minimum (Legacy)"),
        ("Ed25519", "Approved"),
        ("ECDSA P-256+", "Approved"),
        ("ChaCha20", "Approved"),
        ("Weak/Legacy", "NON-COMPLIANT"),
    ]

    row += 1
    algo_start = row
    for algo, status in algorithms:
        ws.cell(row=row, column=1, value=algo)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual count
        ws.cell(row=row, column=3, value=f'=IF(SUM(B{algo_start}:B{row})=0,"0%",ROUND(B{row}/SUM(B${algo_start}:B${row})*100,1)&"%")')
        ws.cell(row=row, column=4, value=status)
        row += 1

    # ---------- 4. STORAGE METHOD DISTRIBUTION ----------
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "STORAGE METHOD DISTRIBUTION"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    storage_headers = ["Storage Method", "Count", "Percentage", "Security Rating"]
    for col_idx, header in enumerate(storage_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    storage_methods = [
        ("HSM", "Highest"),
        ("Cloud KMS", "High"),
        ("TPM", "High"),
        ("Software Keystore", "Medium"),
        ("Vault", "Medium"),
        ("File-based", "LOW RISK"),
        ("Database", "LOW RISK"),
    ]

    row += 1
    storage_start = row
    for method, rating in storage_methods:
        ws.cell(row=row, column=1, value=method)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual count
        ws.cell(row=row, column=3, value=f'=IF(SUM(B{storage_start}:B{row})=0,"0%",ROUND(B{row}/SUM(B${storage_start}:B${row})*100,1)&"%")')
        ws.cell(row=row, column=4, value=rating)
        row += 1

    # ---------- 5. KEY ROTATION COVERAGE ----------
    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "KEY ROTATION COVERAGE"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    rotation_headers = ["Rotation Schedule", "Count", "Percentage", "Compliance Status"]
    for col_idx, header in enumerate(rotation_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    rotation_types = [
        ("Automated", "Optimal"),
        ("Manual (Scheduled)", "Acceptable"),
        ("Never / Not Configured", "NON-COMPLIANT"),
    ]

    row += 1
    for rot_type, status in rotation_types:
        ws.cell(row=row, column=1, value=rot_type)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual/formula
        ws.cell(row=row, column=3).fill = styles["input_cell"]["fill"]  # %
        ws.cell(row=row, column=4, value=status)
        row += 1

    # ---------- 6. CERTIFICATE MANAGEMENT HEALTH ----------
    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "CERTIFICATE MANAGEMENT HEALTH"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    cert_headers = ["Certificate Metric", "Count", "Status"]
    for col_idx, header in enumerate(cert_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    cert_metrics = [
        ("Total Certificates", ""),
        ("Valid Certificates (≤397 days)", ""),
        ("Expired Certificates", "❌"),
        ("Non-Compliant (>397 days)", "❌"),
        ("Auto-Renewal Enabled", ""),
        ("Self-Signed (Non-Compliant)", "❌"),
        ("Certificates Expiring <30 days", "⚠️"),
    ]

    row += 1
    for metric, status in cert_metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual/formula
        ws.cell(row=row, column=3, value=status)
        row += 1

    # ---------- 7. FIPS 140-2 VALIDATION STATUS ----------
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "FIPS 140-2 VALIDATION STATUS"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    fips_headers = ["FIPS Validation Level", "Count", "Percentage", "Requirement Met"]
    for col_idx, header in enumerate(fips_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    fips_levels = [
        ("Level 4", "Highest Security"),
        ("Level 3", "High Security"),
        ("Level 2", "Production Minimum"),
        ("Level 1", "Below Requirement"),
        ("Not Validated", "NON-COMPLIANT"),
    ]

    row += 1
    fips_start = row
    for level, requirement in fips_levels:
        ws.cell(row=row, column=1, value=level)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual/formula
        ws.cell(row=row, column=3, value=f'=IF(SUM(B{fips_start}:B{row})=0,"0%",ROUND(B{row}/SUM(B${fips_start}:B${row})*100,1)&"%")')
        ws.cell(row=row, column=4, value=requirement)
        row += 1

    # ---------- 8. CRITICAL GAPS ----------
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "CRITICAL GAPS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    gap_headers = ["Priority", "Assessment Area", "Gap Description", "Responsible Person", "Target Date", "Status"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=False)
    ws.add_data_validation(dv_priority)

    dv_gap_status = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Closed"', allow_blank=False)
    ws.add_data_validation(dv_gap_status)

    for i in range(8):
        row += 1
        ws.cell(row=row, column=1).fill = styles["input_cell"]["fill"]
        dv_priority.add(ws.cell(row=row, column=1))
        
        for col in range(2, 6):
            c = ws.cell(row=row, column=col)
            c.fill = styles["input_cell"]["fill"]
            c.border = styles["border"]
            c.alignment = styles["input_cell"]["alignment"]
        
        ws.cell(row=row, column=6).fill = styles["input_cell"]["fill"]
        dv_gap_status.add(ws.cell(row=row, column=6))

    # ---------- 9. RISK SUMMARY BY DATA CLASSIFICATION ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "RISK SUMMARY BY DATA CLASSIFICATION"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    risk_headers = ["Data Classification", "Keys/Certs Managed", "Compliant", "Non-Compliant", "Risk Level"]
    for col_idx, header in enumerate(risk_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    classifications = ["Restricted", "Confidential", "Internal", "Public"]
    row += 1
    for classification in classifications:
        ws.cell(row=row, column=1, value=classification)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual count
        ws.cell(row=row, column=3).fill = styles["input_cell"]["fill"]  # Manual count
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]  # Manual count
        ws.cell(row=row, column=5, value=f'=IF(D{row}>0,"HIGH",IF(C{row}/B{row}<0.9,"MEDIUM","LOW"))')
        row += 1

    # ---------- 10. KEY BACKUP COVERAGE ----------
    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "KEY BACKUP COVERAGE"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    backup_headers = ["Backup Status", "Count", "Percentage"]
    for col_idx, header in enumerate(backup_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    backup_statuses = [
        "Backed Up",
        "Not Backed Up",
        "Recovery Tested",
        "Recovery Not Tested",
    ]

    row += 1
    backup_start = row
    for status in backup_statuses:
        ws.cell(row=row, column=1, value=status)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]  # Manual/formula
        ws.cell(row=row, column=3, value=f'=IF(SUM(B{backup_start}:B{row})=0,"0%",ROUND(B{row}/SUM(B${backup_start}:B${row})*100,1)&"%")')
        row += 1

    # ---------- 11. OVERALL ASSESSMENT SUMMARY ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "OVERALL ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "OVERALL KEY MANAGEMENT COMPLIANCE:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f"='Summary Dashboard'!G{total_row}"
    ws[f"B{row}"].font = Font(bold=True, size=12, color="0000FF")

    row += 1
    ws[f"A{row}"] = "Assessment Status:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    dv_assessment = DataValidation(
        type="list",
        formula1='"Excellent (≥95%),Good (90-94%),Adequate (80-89%),Needs Improvement (70-79%),Critical (<70%)"',
        allow_blank=False
    )
    ws.add_data_validation(dv_assessment)
    dv_assessment.add(ws[f"B{row}"])

    row += 2
    ws[f"A{row}"] = "Key Findings Summary:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    ws.merge_cells(f"A{row}:E{row+4}")
    ws[f"A{row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{row}"].border = styles["border"]
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.freeze_panes = "A4"


# ============================================================================
# END OF PART 4
# ============================================================================

# ============================================================================
# SECTION 10: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create evidence register for audit traceability."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID",
        "Assessment Area",
        "Evidence Type",
        "Description",
        "Location/Path",
        "Date Collected",
        "Collected By",
        "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Validation dropdowns
    dv_type = DataValidation(
        type="list",
        formula1='"Key generation log,HSM configuration,KMS architecture,Rotation log,Backup procedure,Recovery test,Certificate inventory,CA audit report,FIPS certificate,Key ceremony record,Access log,Destruction record,Algorithm verification,Third-party audit,Other"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_type)

    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_ver)

    # Data rows (100 rows as per spec)
    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EV-{r-4:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 11: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create approval and sign-off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", "ISMS-IMP-A.8.24.4 - Key Management Assessment"),
        ("Assessment Period", ""),
        ("Overall Compliance Rate", "='Summary Dashboard'!G8"),  # Total row from dashboard
        ("Assessment Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)

        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Status dropdown
    status_cell_row = row - 1
    dv_status = DataValidation(
        type="list", 
        formula1='"Draft,Final,Requires remediation,Re-assessment required"', 
        allow_blank=False
    )
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{status_cell_row}"])

    # ---------- ASSESSMENT COMPLETED BY ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    completion_fields = ["Name", "Role/Title", "Department", "Email", "Date"]
    row += 1
    for field in completion_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # ---------- REVIEWED BY (INFORMATION SECURITY OFFICER) ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (INFORMATION SECURITY OFFICER)"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    review_fields = ["Name", "Date", "Review Notes"]
    row += 1
    for field in review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # ---------- APPROVED BY (CISO) ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    approval_fields = ["Name", "Date", "Approval Decision", "Conditions/Notes"]
    row += 1
    for field in approval_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Approval Decision dropdown (3rd field from approval section start)
    decision_cell_row = row - 2
    dv_dec = DataValidation(
        type="list", 
        formula1='"Approved,Approved with conditions,Rejected"', 
        allow_blank=False
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(ws[f"B{decision_cell_row}"])

    # ---------- NEXT REVIEW DETAILS ----------
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    next_review_fields = ["Next Review Date", "Review Responsible", "Special Considerations"]
    row += 1
    for field in next_review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    print("=" * 78)
    print("ISMS-IMP-A.8.24.4 - Key Management Assessment Generator")
    print("ISO/IEC 27001:2022 Control A.8.24: Use of Cryptography (Key Management)")
    print("=" * 78)

    wb = create_workbook()
    styles = setup_styles()

    print("\n[1/9] Creating Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)

    print("[2/9] Creating 1. Key Generation sheet...")
    create_1_key_generation(wb["1. Key Generation"], styles)

    print("[3/9] Creating 2. Key Storage sheet...")
    create_2_key_storage(wb["2. Key Storage"], styles)

    print("[4/9] Creating 3. Key Rotation sheet...")
    create_3_key_rotation(wb["3. Key Rotation"], styles)

    print("[5/9] Creating 4. Key Backup & Recovery sheet...")
    create_4_key_backup_recovery(wb["4. Key Backup & Recovery"], styles)

    print("[6/9] Creating 5. Certificate Management sheet...")
    create_5_certificate_management(wb["5. Certificate Management"], styles)

    print("[7/9] Creating Summary Dashboard sheet...")
    create_summary_dashboard(wb["Summary Dashboard"], styles)

    print("[8/9] Creating Evidence Register sheet...")
    create_evidence_register(wb["Evidence Register"], styles)

    print("[9/9] Creating Approval Sign-Off sheet...")
    create_approval_signoff(wb["Approval Sign-Off"], styles)

    filename = f"ISMS-IMP-A.8.24.4_Key_Management_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    print(f"\n✅ SUCCESS: {filename}")
    print("\nWorkbook Structure:")
    print("  • Instructions & Legend - Document info and guidance")
    print("  • 5 Assessment Sheets:")
    print("    1. Key Generation (14 checklist items + 1 reference table)")
    print("    2. Key Storage (17 checklist items + 1 reference table)")
    print("    3. Key Rotation (16 checklist items + 1 reference table)")
    print("    4. Key Backup & Recovery (18 checklist items + 1 reference table)")
    print("    5. Certificate Management (20 checklist items + 2 reference tables)")
    print("  • Summary Dashboard - 11 comprehensive analysis sections:")
    print("    - Compliance summary by key management area")
    print("    - Key management maturity assessment")
    print("    - Algorithm distribution analysis")
    print("    - Storage method distribution with security ratings")
    print("    - Key rotation coverage analysis")
    print("    - Certificate management health metrics")
    print("    - FIPS 140-2 validation status")
    print("    - Critical gaps requiring immediate attention (8 rows)")
    print("    - Risk summary by data classification")
    print("    - Key backup coverage metrics")
    print("    - Overall assessment summary with maturity level")
    print("  • Evidence Register - 100 evidence entry rows")
    print("  • Approval Sign-Off - Complete workflow with 3 sign-off levels")
    
    print("\nColumn Structure:")
    print("  • Base columns (A-Q): 17 standard columns across all sheets")
    print("  • Extended columns (R-X):")
    print("    - Key Generation: +5 columns (RNG, location, ceremony, entropy, audit)")
    print("    - Key Storage: +6 columns (HSM/KMS, FIPS level, access control, etc.)")
    print("    - Key Rotation: +6 columns (schedule, dates, automation, testing, etc.)")
    print("    - Key Backup & Recovery: +7 columns (method, encryption, location, RTO, etc.)")
    print("    - Certificate Management: +7 columns (type, CA, validity, renewal, etc.)")
    
    print("\nNext steps:")
    print("  1) Complete document information in Instructions & Legend")
    print("  2) Fill yellow cells in each assessment sheet (1–5)")
    print("  3) Check compliance checklists per key management area:")
    print("     - Key Generation: Approved algorithms, CSRNG, FIPS validation")
    print("     - Key Storage: HSM/KMS usage, FIPS 140-2 Level 2+, access controls")
    print("     - Key Rotation: Rotation schedules, automation, testing")
    print("     - Key Backup & Recovery: Backup procedures, recovery testing, escrow")
    print("     - Certificate Management: CA trust, validity periods, revocation")
    print("  4) Review reference tables (approved algorithms, storage solutions, etc.)")
    print("  5) Document exceptions/deviations as needed")
    print("  6) Maintain Evidence Register entries")
    print("  7) Review Summary Dashboard:")
    print("     - Overall compliance by key management domain")
    print("     - Maturity assessment per area")
    print("     - Algorithm and storage method distribution")
    print("     - Key rotation and backup coverage")
    print("     - Certificate health and FIPS validation status")
    print("     - Critical gaps and risk summary")
    print("  8) Complete multi-level approval sign-off workflow")
    print("\n" + "=" * 78)


if __name__ == "__main__":
    main()


# ============================================================================
# END OF SCRIPT - COMPLETE KEY MANAGEMENT ASSESSMENT GENERATOR
# ============================================================================