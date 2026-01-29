#!/usr/bin/env python3
"""
ISMS-IMP-A.8.23.1 - Filtering Infrastructure Assessment Excel Generator
ISO/IEC 27001:2022 Control A.8.23 (Web Filtering - Infrastructure)

Requirements:
    sudo apt install python3-openpyxl
    
Usage:
    python3 generate_a823_1_filtering_infrastructure.py

"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches ISMS-IMP-A.8.23.1 specification
    sheets = [
        "Instructions & Legend",
        "Solution_Details_Template",
        "Technology_Comparison",
        "Capability_Requirements",
        "Integration_Architecture",
        "Licensing_Support",
        "Performance_Metrics",
        "Gap_Analysis",
        "Evidence_Register",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    CRITICAL: Each style object must be created NEW for each cell to avoid
    Excel repair warnings from shared Border/Font/Fill objects.
    This function returns TEMPLATES, not reusable objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Return style TEMPLATES (dictionaries), not objects
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_deployed": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_notdeployed": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_planned": {
            "fill": PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        },
        "gap_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "gap_high": {
            "fill": PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),
        },
        "gap_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "gap_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell.
    Creates NEW style objects to avoid shared object warnings.
    """
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ============================================================================
# SECTION 2: DATA VALIDATIONS & COLUMN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """
    Create data validation objects for standard dropdowns.
    These are added to worksheet once, then applied to multiple cells.
    """
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_partial': DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Unknown"',
            allow_blank=False
        ),
        'yes_no_planned_na': DataValidation(
            type="list",
            formula1='"Yes,No,Planned,N/A"',
            allow_blank=False
        ),
        'deployment_status': DataValidation(
            type="list",
            formula1='"✅ Deployed,⚠️ Partial,❌ Not Deployed,🔄 Planned,N/A"',
            allow_blank=False
        ),
        'deployment_model': DataValidation(
            type="list",
            formula1='"On-Premises Appliance,Virtual Appliance,Cloud-Based SaaS,Hybrid,DNS-Based,Proxy-Based,Other"',
            allow_blank=False
        ),
        'primary_purpose': DataValidation(
            type="list",
            formula1='"Threat Protection,Category Filtering,Compliance (CIPA/etc),Bandwidth Management,Hybrid/All"',
            allow_blank=False
        ),
        'license_type': DataValidation(
            type="list",
            formula1='"Perpetual,Subscription,Pay-per-use,Open Source,Unknown"',
            allow_blank=False
        ),
        'support_status': DataValidation(
            type="list",
            formula1='"Yes,No,Expired"',
            allow_blank=False
        ),
        'support_level': DataValidation(
            type="list",
            formula1='"24/7,Business Hours,Community,None"',
            allow_blank=False
        ),
        'update_schedule': DataValidation(
            type="list",
            formula1='"Automatic,Manual-Monthly,Manual-Quarterly,Ad-hoc"',
            allow_blank=False
        ),
        'update_frequency': DataValidation(
            type="list",
            formula1='"Real-time,Daily,Weekly,Monthly,Unknown"',
            allow_blank=False
        ),
        'rating': DataValidation(
            type="list",
            formula1='"Low,Medium,High,Unknown"',
            allow_blank=False
        ),
        'quality_rating': DataValidation(
            type="list",
            formula1='"Excellent,Good,Adequate,Poor"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'gap_status': DataValidation(
            type="list",
            formula1='"Open,In Progress,Resolved,Closed"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Config File,Screenshot,Report,License,Contract,Log,Diagram,Policy,Other"',
            allow_blank=False
        ),
        'verification_status': DataValidation(
            type="list",
            formula1='"Verified,Pending,Not Verified"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1='"Draft,Final,Requires Remediation,Re-assessment Required"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected"',
            allow_blank=False
        ),
        'integration_type_ad': DataValidation(
            type="list",
            formula1='"Native,LDAP,RADIUS,SAML,None"',
            allow_blank=False
        ),
        'integration_type_siem': DataValidation(
            type="list",
            formula1='"Syslog,API,Agent,None"',
            allow_blank=False
        ),
        'integration_type_proxy': DataValidation(
            type="list",
            formula1='"Integrated,Chained,Standalone"',
            allow_blank=False
        ),
        'incident_type': DataValidation(
            type="list",
            formula1='"Outage,Performance,False Positive,Bypass,Other"',
            allow_blank=False
        ),
        'incident_severity': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
    }

    # Add all validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)

    return validations


def get_solution_details_fields():
    """
    Return field definitions for Solution_Details_Template sheet.
    Format: [(field_name, field_type, validation_key, notes)]
    """
    fields = [
        # Solution Identification Section
        ("Solution Name/Description", "text", None, "What YOU call this solution"),
        ("Vendor/Provider", "text", None, "Customer fills in (e.g., Sophos, Fortigate, Zscaler, etc.)"),
        ("Product/Service Name", "text", None, "Customer fills in specific product"),
        ("Version/Release", "text", None, "Current version deployed"),
        ("Deployment Model", "dropdown", "deployment_model", None),
        ("Deployment Date", "date", None, "When was it first deployed?"),
        ("Deployment Location(s)", "text", None, "Network segments where deployed"),
        ("Primary Purpose", "dropdown", "primary_purpose", None),
        ("Status", "dropdown", "deployment_status", None),
        ("Scope of Coverage", "text", None, "What does this protect?"),
        ("User Count", "number", None, "Approximate number of protected users"),
        ("Integration Points", "text", None, "What does it integrate with?"),
    ]
    return fields


def get_capability_assessments():
    """
    Return capability assessment items for Solution_Details_Template.
    Format: [(category, capability, validation_key)]
    """
    capabilities = [
        # Threat Protection
        ("THREAT PROTECTION", "Blocks known malicious URLs", "yes_no_partial"),
        ("THREAT PROTECTION", "Blocks phishing sites", "yes_no_partial"),
        ("THREAT PROTECTION", "Malware download prevention", "yes_no_partial"),
        ("THREAT PROTECTION", "Ransomware protection", "yes_no_partial"),
        ("THREAT PROTECTION", "Exploit prevention", "yes_no_partial"),
        ("THREAT PROTECTION", "Zero-day threat protection", "yes_no_partial"),
        
        # URL Categorization
        ("URL CATEGORIZATION", "Supports URL categorization", "yes_no"),
        ("URL CATEGORIZATION", "Number of categories available", "number"),
        ("URL CATEGORIZATION", "Category database update frequency", "update_frequency"),
        
        # Content Analysis
        ("CONTENT ANALYSIS", "HTTPS/SSL inspection capable", "yes_no_partial"),
        ("CONTENT ANALYSIS", "HTTPS inspection enabled?", "yes_no_planned_na"),
        ("CONTENT ANALYSIS", "File type filtering", "yes_no"),
        ("CONTENT ANALYSIS", "Data Loss Prevention (DLP)", "yes_no"),
        
        # Policy Enforcement
        ("POLICY ENFORCEMENT", "User-based policies", "yes_no"),
        ("POLICY ENFORCEMENT", "Group-based policies", "yes_no"),
        ("POLICY ENFORCEMENT", "Time-based policies", "yes_no"),
        ("POLICY ENFORCEMENT", "Location-based policies", "yes_no"),
        
        # Logging & Monitoring
        ("LOGGING & MONITORING", "Generates access logs", "yes_no"),
        ("LOGGING & MONITORING", "Generates block/alert logs", "yes_no"),
        ("LOGGING & MONITORING", "Real-time alerting", "yes_no"),
        ("LOGGING & MONITORING", "Reporting/dashboards", "yes_no"),
        ("LOGGING & MONITORING", "Log retention capability", "text"),
        
        # Administration
        ("ADMINISTRATION", "Centralized management", "yes_no_na"),
        ("ADMINISTRATION", "Multi-admin support", "yes_no"),
        ("ADMINISTRATION", "Role-based admin access", "yes_no"),
        ("ADMINISTRATION", "Configuration backup/restore", "yes_no"),
        ("ADMINISTRATION", "Change audit logging", "yes_no"),
        
        # Integration
        ("INTEGRATION", "Active Directory integration", "yes_no_na"),
        ("INTEGRATION", "SIEM integration", "yes_no"),
        ("INTEGRATION", "Threat intelligence feeds", "yes_no"),
        ("INTEGRATION", "API access for automation", "yes_no_partial"),
    ]
    return capabilities


def get_licensing_fields():
    """Return licensing and support fields for Solution_Details_Template."""
    fields = [
        ("License Type", "dropdown", "license_type"),
        ("License Expiration Date", "date", None),
        ("Licensed User/Device Count", "number", None),
        ("Support Contract Active?", "dropdown", "support_status"),
        ("Support Level", "dropdown", "support_level"),
        ("Support Expiration Date", "date", None),
        ("Update/Patch Schedule", "dropdown", "update_schedule"),
        ("Last Update Applied", "date", None),
        ("Threat Database Version", "text", None),
        ("Threat Database Last Updated", "date", None),
        ("Annual License Cost", "number", None),
        ("Annual Support Cost", "number", None),
    ]
    return fields


def get_performance_fields():
    """Return performance metrics fields for Solution_Details_Template."""
    fields = [
        ("Uptime SLA (if applicable)", "text", None),
        ("Actual uptime (last quarter)", "text", None),
        ("Average latency impact", "text", None),
        ("False positive rate", "dropdown", "rating"),
        ("False negative rate", "dropdown", "rating"),
        ("Incident count (last 12 months)", "number", None),
        ("Performance monitoring enabled?", "dropdown", "yes_no"),
        ("Capacity utilization", "text", None),
        ("Scalability", "dropdown", "quality_rating"),
        ("Redundancy/HA configured?", "dropdown", "yes_no_na"),
    ]
    return fields

# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet with comprehensive guidance."""
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "ISMS-IMP-A.8.23.1 – Filtering Infrastructure Assessment"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "ISO/IEC 27001:2022 - Control A.8.23: Web Filtering"
    apply_style(ws["A2"], styles["subheader"])

    # Document Information Block
    row = 4
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    
    doc_info = [
        ("Document ID:", "ISMS-IMP-A.8.23.1"),
        ("Assessment Area:", "Web Filtering Infrastructure"),
        ("Related Policy:", "ISMS-POL-A.8.23-S2.1, S2.2"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Quarterly"),
    ]
    
    row += 1
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if "USER INPUT" in value:
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    # How to Use This Workbook
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    instructions = [
        "1. Complete the Solution_Details_Template for EACH filtering solution you have deployed",
        "2. Fill in YOUR specific vendor/product names (this workbook is vendor-agnostic)",
        "3. Use dropdown menus for standardized capability assessments",
        "4. Document licensing, support contracts, and performance metrics",
        "5. Complete the Technology_Comparison sheet if you have multiple solutions",
        "6. Assess capabilities against policy requirements in Capability_Requirements sheet",
        "7. Identify gaps in the Gap_Analysis sheet",
        "8. Maintain the Evidence Register for audit traceability",
        "9. Obtain final approval and sign-off",
    ]

    row += 1
    for instruction in instructions:
        ws[f"A{row}"] = instruction
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1

    # Status Legend
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "STATUS LEGEND"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    legend_headers = ["Symbol", "Status", "Description", "Color Code"]
    for col_idx, header in enumerate(legend_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    legend_data = [
        ("✅", "Deployed", "Solution deployed and operational", "Green"),
        ("⚠️", "Partial", "Partial deployment or limited functionality", "Yellow"),
        ("❌", "Not Deployed", "Solution not deployed", "Red"),
        ("🔄", "Planned", "Deployment planned with target date", "Blue"),
        ("N/A", "Not Applicable", "Not applicable to this environment", "Gray"),
    ]

    row += 1
    for symbol, status, desc, color in legend_data:
        ws[f"A{row}"] = symbol
        ws[f"B{row}"] = status
        ws[f"C{row}"] = desc
        ws[f"D{row}"] = color
        
        # Apply color coding
        if color == "Green":
            ws[f"D{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif color == "Yellow":
            ws[f"D{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif color == "Red":
            ws[f"D{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif color == "Blue":
            ws[f"D{row}"].fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        
        row += 1

    # Acceptable Evidence
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (Examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    evidence_examples = [
        "✓ Network diagrams showing filtering placement",
        "✓ Configuration backups/exports (sanitized)",
        "✓ License agreements and support contracts",
        "✓ Performance monitoring dashboards (screenshots)",
        "✓ Vendor capability documentation",
        "✓ Integration architecture diagrams",
        "✓ Change management records for updates/patches",
        "✓ Incident response logs (filtering-related)",
        "✓ Administrative access logs",
        "✓ Compliance reports from the filtering solution",
        "✓ Threat intelligence feed configurations",
        "✓ HTTPS inspection certificates/policies",
    ]

    row += 1
    for evidence in evidence_examples:
        ws[f"A{row}"] = evidence
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    ws.freeze_panes = "A4"

# ============================================================================
# SECTION 4: SOLUTION_DETAILS_TEMPLATE SHEET
# ============================================================================

def create_solution_details_template(ws, styles):
    """
    Create Solution_Details_Template sheet.
    This is the core assessment template - copy for each filtering solution.
    """
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:D1")
    ws["A1"] = "WEB FILTERING SOLUTION DETAILS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    ws["A2"] = "Complete one copy of this template for EACH filtering solution deployed"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== SOLUTION IDENTIFICATION SECTION ====================
    row = 4
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "SOLUTION IDENTIFICATION"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Field"
    ws[f"B{row}"] = "Value"
    ws[f"C{row}"] = "Evidence Location"
    ws[f"D{row}"] = "Notes"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Solution identification fields
    id_fields = get_solution_details_fields()
    row += 1
    start_id_row = row
    
    for field_name, field_type, validation_key, notes in id_fields:
        ws[f"A{row}"] = field_name
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        if notes:
            ws[f"D{row}"] = notes
            ws[f"D{row}"].font = Font(italic=True, size=9)
        
        # Apply validation if specified
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])
        
        row += 1

    # ==================== CAPABILITY ASSESSMENT SECTION ====================
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "CAPABILITY ASSESSMENT (What can this solution do?)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Capability"
    ws[f"B{row}"] = "Assessment"
    ws[f"C{row}"] = "Evidence Location"
    ws[f"D{row}"] = "Notes"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Capability assessments
    capabilities = get_capability_assessments()
    row += 1
    current_category = None
    
    for category, capability, validation_key in capabilities:
        # Insert category header if new category
        if category != current_category:
            ws[f"A{row}"] = f"** {category} **"
            ws[f"A{row}"].font = Font(bold=True, size=10, color="4472C4")
            ws.merge_cells(f"A{row}:D{row}")
            row += 1
            current_category = category
        
        # Capability row
        ws[f"A{row}"] = capability
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Apply validation
        if validation_key == "number":
            # Number field - no dropdown
            pass
        elif validation_key == "text":
            # Text field - no dropdown
            pass
        elif validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])
        
        row += 1

    # ==================== LICENSING & SUPPORT SECTION ====================
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "LICENSING & SUPPORT"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Field"
    ws[f"B{row}"] = "Value"
    ws[f"C{row}"] = "Evidence"
    ws[f"D{row}"] = "Notes"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Licensing fields
    lic_fields = get_licensing_fields()
    row += 1
    
    for field_name, field_type, validation_key in lic_fields:
        ws[f"A{row}"] = field_name
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Apply validation if specified
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])
        
        row += 1

    # ==================== PERFORMANCE & RELIABILITY SECTION ====================
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "PERFORMANCE & RELIABILITY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"B{row}"] = "Value"
    ws[f"C{row}"] = "Evidence"
    ws[f"D{row}"] = "Notes"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Performance fields
    perf_fields = get_performance_fields()
    row += 1
    
    for field_name, field_type, validation_key in perf_fields:
        ws[f"A{row}"] = field_name
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Apply validation if specified
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"B{row}"])
        
        row += 1

    # ==================== GAP IDENTIFICATION SECTION ====================
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "GAP IDENTIFICATION"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Gap Description"
    ws[f"B{row}"] = "Severity"
    ws[f"C{row}"] = "Target Date"
    ws[f"D{row}"] = "Responsible Person"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # 8 gap rows
    row += 1
    for i in range(8):
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        validations['risk_level'].add(ws[f"B{row}"])
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30

    ws.freeze_panes = "A4"

# ============================================================================
# SECTION 5: TECHNOLOGY_COMPARISON SHEET
# ============================================================================

def create_technology_comparison(ws, styles):
    """
    Create Technology_Comparison sheet for side-by-side comparison.
    Supports up to 4 solutions (can be extended).
    """
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "TECHNOLOGY COMPARISON MATRIX"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "Compare capabilities across all deployed filtering solutions"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = ["Capability", "Solution 1", "Solution 2", "Solution 3", "Solution 4", "Best Coverage"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Comparison categories and items
    comparison_items = [
        ("DEPLOYMENT", [
            "Deployment Model",
            "Deployment Location",
            "User Count",
            "Status",
        ]),
        ("THREAT PROTECTION", [
            "Malicious URL blocking",
            "Phishing protection",
            "Malware prevention",
            "Ransomware protection",
            "Zero-day protection",
        ]),
        ("CONTENT INSPECTION", [
            "HTTPS inspection capable",
            "HTTPS inspection enabled",
            "File type filtering",
            "DLP capabilities",
        ]),
        ("POLICY CAPABILITIES", [
            "User-based policies",
            "Group-based policies",
            "Time-based policies",
            "Location-based policies",
        ]),
        ("LOGGING & MONITORING", [
            "Access logging",
            "Block/alert logging",
            "Real-time alerting",
            "Reporting dashboards",
        ]),
        ("INTEGRATION", [
            "Active Directory",
            "SIEM integration",
            "Threat intelligence feeds",
            "API access",
        ]),
        ("SUPPORT & LICENSING", [
            "License type",
            "Support level",
            "Support quality rating",
            "Annual cost",
        ]),
    ]

    row += 1
    for category, items in comparison_items:
        # Category header
        ws[f"A{row}"] = f"** {category} **"
        ws[f"A{row}"].font = Font(bold=True, size=10, color="4472C4")
        ws.merge_cells(f"A{row}:F{row}")
        row += 1
        
        # Items
        for item in items:
            ws[f"A{row}"] = item
            ws[f"A{row}"].font = Font(bold=True, size=9)
            
            # Solution columns (B-E) - customer fills in
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")
            
            # Best coverage column (F) - dropdown
            ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            # Create validation for "which solution is best" dropdown
            best_dv = DataValidation(
                type="list",
                formula1='"Solution 1,Solution 2,Solution 3,Solution 4,Equal,None"',
                allow_blank=True
            )
            ws.add_data_validation(best_dv)
            best_dv.add(ws[f"F{row}"])
            
            row += 1

    # Notes section
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "COMPARISON NOTES & OBSERVATIONS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws.merge_cells(f"A{row}:F{row+10}")
    ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18

    ws.freeze_panes = "A5"

# ============================================================================
# SECTION 6: CAPABILITY_REQUIREMENTS SHEET
# ============================================================================

def create_capability_requirements(ws, styles):
    """
    Create Capability_Requirements sheet.
    Maps policy requirements to deployed solutions.
    """
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "CAPABILITY REQUIREMENTS vs. DEPLOYED SOLUTIONS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "Verify policy compliance across all deployed filtering technologies"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = ["Requirement ID", "Policy Requirement", "Met by Solution(s)", "Status", "Gap?", "Evidence"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Policy requirements (based on ISMS-POL-A.8.23)
    requirements = [
        ("REQ-001", "SHALL block known malicious URLs"),
        ("REQ-002", "SHALL block phishing sites"),
        ("REQ-003", "SHALL prevent malware downloads"),
        ("REQ-004", "SHOULD block ransomware delivery sites"),
        ("REQ-005", "SHOULD support HTTPS inspection"),
        ("REQ-006", "SHALL log all filtering events"),
        ("REQ-007", "SHALL retain logs for ≥90 days"),
        ("REQ-008", "MUST support threat feed updates"),
        ("REQ-009", "SHOULD support URL categorization"),
        ("REQ-010", "SHOULD support user-based policies"),
        ("REQ-011", "SHOULD support group-based policies"),
        ("REQ-012", "SHALL provide administrative audit logs"),
        ("REQ-013", "MUST have configuration backup capability"),
        ("REQ-014", "SHOULD integrate with SIEM"),
        ("REQ-015", "SHOULD integrate with Active Directory"),
        ("REQ-016", "SHALL support policy exceptions with approval"),
        ("REQ-017", "MUST provide real-time or near-real-time blocking"),
        ("REQ-018", "SHOULD provide reporting dashboards"),
        ("REQ-019", "SHALL support incident response procedures"),
        ("REQ-020", "MUST be manageable by authorized admins only"),
        ("REQ-021", "SHOULD support role-based admin access"),
        ("REQ-022", "SHALL have documented change management"),
        ("REQ-023", "MUST have valid support contracts"),
        ("REQ-024", "SHOULD have ≥99% uptime SLA"),
        ("REQ-025", "SHALL minimize false positives"),
        ("REQ-026", "SHOULD support time-based policies (optional)"),
        ("REQ-027", "SHOULD support DLP integration (optional)"),
        ("REQ-028", "SHALL comply with acceptable use policy"),
        ("REQ-029", "MUST protect against zero-day threats (best effort)"),
        ("REQ-030", "SHOULD support API access for automation"),
    ]

    row += 1
    for req_id, req_text in requirements:
        ws[f"A{row}"] = req_id
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        ws[f"B{row}"] = req_text
        ws[f"B{row}"].alignment = Alignment(wrap_text=True)
        
        # Met by Solution(s) - customer fills in
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Status dropdown
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        status_dv = DataValidation(
            type="list",
            formula1='"✅ Met,⚠️ Partial,❌ Not Met,N/A"',
            allow_blank=False
        )
        ws.add_data_validation(status_dv)
        status_dv.add(ws[f"D{row}"])
        
        # Gap? dropdown
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['yes_no'].add(ws[f"E{row}"])
        
        # Evidence
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # Summary metrics section
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "COMPLIANCE SUMMARY METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    summary_items = [
        ("Total Requirements:", "=COUNTA(A5:A34)"),
        ("Requirements Met (✅):", f'=COUNTIF(D5:D34,"✅ Met")'),
        ("Partially Met (⚠️):", f'=COUNTIF(D5:D34,"⚠️ Partial")'),
        ("Not Met (❌):", f'=COUNTIF(D5:D34,"❌ Not Met")'),
        ("N/A:", f'=COUNTIF(D5:D34,"N/A")'),
        ("Compliance Rate:", '=(COUNTIF(D5:D34,"✅ Met")/COUNTA(A5:A34))*100&"%"'),
    ]

    for label, formula in summary_items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, size=11, color="0000FF")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 30

    ws.freeze_panes = "A5"

# ============================================================================
# SECTION 7: INTEGRATION_ARCHITECTURE SHEET
# ============================================================================

def create_integration_architecture(ws, styles):
    """
    Create Integration_Architecture sheet.
    Documents how web filtering integrates with existing infrastructure.
    """
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "INTEGRATION ARCHITECTURE"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    ws["A2"] = "Document integration points with existing infrastructure"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = ["Integration Point", "Solution Name", "Integration Type", "Status", "Evidence"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Integration points with specific validation types
    integrations = [
        ("Active Directory", "integration_type_ad"),
        ("SIEM", "integration_type_siem"),
        ("Proxy Server", "integration_type_proxy"),
        ("DNS", None),  # Custom validation below
        ("Firewall", None),
        ("Endpoint Security", None),
        ("Email Gateway", None),
        ("Threat Intelligence", None),
        ("IAM/SSO", None),
        ("DLP", None),
    ]

    row += 1
    for integration_point, validation_key in integrations:
        ws[f"A{row}"] = integration_point
        ws[f"A{row}"].font = Font(bold=True)
        
        # Solution Name (customer fills in)
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Integration Type (dropdown)
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        if validation_key and validation_key in validations:
            validations[validation_key].add(ws[f"C{row}"])
        else:
            # Generic integration type for others
            generic_int_dv = DataValidation(
                type="list",
                formula1='"Native,API,Agent,Syslog,LDAP,SAML,Manual,None"',
                allow_blank=False
            )
            ws.add_data_validation(generic_int_dv)
            generic_int_dv.add(ws[f"C{row}"])
        
        # Status
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        status_dv = DataValidation(
            type="list",
            formula1='"✅ Integrated,⚠️ Partial,❌ Not Integrated,🔄 Planned"',
            allow_blank=False
        )
        ws.add_data_validation(status_dv)
        status_dv.add(ws[f"D{row}"])
        
        # Evidence
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # Network Architecture Notes section
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NETWORK ARCHITECTURE NOTES"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    note_sections = [
        "Where is filtering applied? (perimeter, endpoint, cloud, hybrid?)",
        "Traffic flow description:",
        "Bypass scenarios (if any):",
        "Redundancy/failover configuration:",
    ]

    for note_section in note_sections:
        ws[f"A{row}"] = note_section
        ws[f"A{row}"].font = Font(bold=True, size=10)
        row += 1
        
        ws.merge_cells(f"A{row}:E{row+2}")
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 3

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 35

    ws.freeze_panes = "A5"

# ============================================================================
# SECTION 8: LICENSING_SUPPORT SHEET
# ============================================================================

def create_licensing_support(ws, styles):
    """
    Create Licensing_Support sheet.
    Centralized tracking of licenses and support contracts.
    """
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "LICENSING & SUPPORT STATUS"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "Maintain awareness of expiration dates and renewal requirements"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== LICENSE REGISTRY ====================
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "LICENSE REGISTRY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    headers = ["Solution Name", "License Type", "User/Device Count", "Expiration Date", 
               "Days Until Expiry", "Status", "Renewal Process Owner"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # License rows (10 rows)
    row += 1
    license_start_row = row
    for i in range(10):
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Days Until Expiry formula (assuming today is in a hidden cell or use NOW())
        ws[f"E{row}"] = f'=IF(D{row}="","",D{row}-TODAY())'
        ws[f"E{row}"].number_format = '0'
        
        # Status dropdown
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        status_lic_dv = DataValidation(
            type="list",
            formula1='"Active,Expiring Soon,Expired"',
            allow_blank=True
        )
        ws.add_data_validation(status_lic_dv)
        status_lic_dv.add(ws[f"F{row}"])
        
        # Owner
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Conditional formatting for Days Until Expiry
        # <30 days: Red, 30-90: Yellow, >90: Green
        # Note: Conditional formatting via openpyxl is complex, document in instructions instead
        
        row += 1

    # ==================== SUPPORT CONTRACT REGISTRY ====================
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "SUPPORT CONTRACT REGISTRY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    support_headers = ["Solution Name", "Support Level", "Expiration Date", 
                      "Days Until Expiry", "Last Support Ticket", "Support Quality Rating"]
    for col_idx, header in enumerate(support_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Support contract rows (10 rows)
    row += 1
    for i in range(10):
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Days Until Expiry formula
        ws[f"D{row}"] = f'=IF(C{row}="","",C{row}-TODAY())'
        ws[f"D{row}"].number_format = '0'
        
        # Last Support Ticket date
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Quality Rating
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['quality_rating'].add(ws[f"F{row}"])
        
        row += 1

    # ==================== UPDATE/PATCH STATUS ====================
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "UPDATE/PATCH STATUS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    update_headers = ["Solution Name", "Update Schedule", "Last Update Date", 
                     "Days Since Update", "Threat DB Version", "DB Last Updated"]
    for col_idx, header in enumerate(update_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Update status rows (10 rows)
    row += 1
    for i in range(10):
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Days Since Update formula
        ws[f"D{row}"] = f'=IF(C{row}="","",TODAY()-C{row})'
        ws[f"D{row}"].number_format = '0'
        
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # Alert Conditions note
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "⚠️ ALERT CONDITIONS: >90 days since update = Review required | Threat DB >7 days old = Review required"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="C00000")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 25

    ws.freeze_panes = "A5"

# ============================================================================
# SECTION 9: PERFORMANCE_METRICS SHEET
# ============================================================================

def create_performance_metrics(ws, styles):
    """
    Create Performance_Metrics sheet.
    Track uptime, latency, false positives, and incidents.
    """
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "PERFORMANCE METRICS & RELIABILITY TRACKING"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Monitor uptime, latency, and incident trends"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== UPTIME TRACKING ====================
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "UPTIME TRACKING"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    uptime_headers = ["Solution Name", "SLA Target", "Q1 Actual", "Q2 Actual", 
                     "Q3 Actual", "Q4 Actual", "Annual Average", "Met SLA?"]
    for col_idx, header in enumerate(uptime_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Uptime rows (5 solutions)
    row += 1
    for i in range(5):
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Annual Average formula
        ws[f"G{row}"] = f'=IF(COUNT(C{row}:F{row})=0,"",AVERAGE(C{row}:F{row}))'
        ws[f"G{row}"].number_format = '0.00"%"'
        
        # Met SLA? dropdown
        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['yes_no'].add(ws[f"H{row}"])
        
        row += 1

    # ==================== LATENCY IMPACT ====================
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "LATENCY IMPACT"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    latency_headers = ["Solution Name", "Baseline Latency (no filter)", "With Filter", "Impact (ms)", "Acceptable?"]
    for col_idx, header in enumerate(latency_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Latency rows (5 solutions)
    row += 1
    for i in range(5):
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Impact calculation
        ws[f"D{row}"] = f'=IF(OR(B{row}="",C{row}=""),"",C{row}-B{row})'
        ws[f"D{row}"].number_format = '0.0'
        
        # Acceptable?
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['yes_no'].add(ws[f"E{row}"])
        
        row += 1

    # ==================== FALSE POSITIVE/NEGATIVE TRACKING ====================
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "FALSE POSITIVE/NEGATIVE TRACKING"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    fp_headers = ["Month", "False Positives", "False Negatives", "User Complaints", "Remediation Actions"]
    for col_idx, header in enumerate(fp_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Monthly tracking rows (12 months)
    row += 1
    current_year = datetime.now().year
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    
    for month in months:
        ws[f"A{row}"] = f"{current_year}-{month}"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # ==================== INCIDENT LOG ====================
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "INCIDENT LOG"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    incident_headers = ["Date", "Solution", "Incident Type", "Severity", "Duration", "Root Cause", "Resolution"]
    for col_idx, header in enumerate(incident_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Incident rows (20 rows)
    row += 1
    for i in range(20):
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Incident Type dropdown
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['incident_type'].add(ws[f"C{row}"])
        
        # Severity dropdown
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['incident_severity'].add(ws[f"D{row}"])
        
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 12

    ws.freeze_panes = "A5"

# ============================================================================
# SECTION 10: GAP_ANALYSIS SHEET
# ============================================================================

def create_gap_analysis(ws, styles):
    """
    Create Gap_Analysis sheet.
    Consolidated gap identification and remediation tracking.
    """
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:K1")
    ws["A1"] = "GAP ANALYSIS & REMEDIATION ROADMAP"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = "Identify deficiencies and track remediation progress"
    apply_style(ws["A2"], styles["subheader"])

    # ==================== GAP REGISTER ====================
    row = 4
    headers = [
        "Gap ID",
        "Gap Description", 
        "Affected Solution(s)",
        "Policy Requirement",
        "Risk Level",
        "Impact",
        "Remediation Plan",
        "Owner",
        "Target Date",
        "Status",
        "Budget Required"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Gap tracking rows (40 rows)
    row += 1
    gap_start_row = row
    
    for i in range(1, 41):
        # Gap ID (auto-generate)
        ws[f"A{row}"] = f"GAP-{i:03d}"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # Gap Description
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Affected Solution(s)
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Policy Requirement
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Risk Level dropdown
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['risk_level'].add(ws[f"E{row}"])
        
        # Impact
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Remediation Plan
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Owner
        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Target Date
        ws[f"I{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Status dropdown
        ws[f"J{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['gap_status'].add(ws[f"J{row}"])
        
        # Budget Required
        ws[f"K{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['yes_no'].add(ws[f"K{row}"])
        
        row += 1
    
    gap_end_row = row - 1

    # ==================== GAP SUMMARY METRICS ====================
    row += 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "GAP SUMMARY METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Risk Level"
    ws[f"B{row}"] = "Count"
    ws[f"C{row}"] = "% of Total"
    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Risk level summary
    row += 1
    risk_levels = ["Critical", "High", "Medium", "Low"]
    for risk in risk_levels:
        ws[f"A{row}"] = risk
        ws[f"A{row}"].font = Font(bold=True)
        
        # Count formula
        ws[f"B{row}"] = f'=COUNTIF(E{gap_start_row}:E{gap_end_row},"{risk}")'
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        
        # Percentage formula
        ws[f"C{row}"] = f'=IF(B{row}=0,0,B{row}/COUNTA(E{gap_start_row}:E{gap_end_row})*100)&"%"'
        ws[f"C{row}"].font = Font(bold=True, color="0000FF")
        
        # Color code by risk level
        if risk == "Critical":
            ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")
        elif risk == "High":
            ws[f"A{row}"].fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
        elif risk == "Medium":
            ws[f"A{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif risk == "Low":
            ws[f"A{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        row += 1

    # Overall summary
    row += 1
    ws[f"A{row}"] = "Total Gaps Identified:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTA(B{gap_start_row}:B{gap_end_row})'
    ws[f"B{row}"].font = Font(bold=True, size=11, color="C00000")

    row += 1
    ws[f"A{row}"] = "Gaps Resolved:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=COUNTIF(J{gap_start_row}:J{gap_end_row},"Resolved")+COUNTIF(J{gap_start_row}:J{gap_end_row},"Closed")'
    ws[f"B{row}"].font = Font(bold=True, size=11, color="008000")

    row += 1
    ws[f"A{row}"] = "Resolution Rate:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f'=IF(COUNTA(B{gap_start_row}:B{gap_end_row})=0,0,B{row-1}/B{row-2}*100)&"%"'
    ws[f"B{row}"].font = Font(bold=True, size=11, color="0000FF")

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 35
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 15

    ws.freeze_panes = "A5"

# ============================================================================
# SECTION 11: EVIDENCE_REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """
    Create Evidence_Register sheet.
    Centralized evidence repository with 100 rows.
    """
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence supporting this assessment"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        "Evidence ID",
        "Evidence Type",
        "Description",
        "Related Sheet/Row",
        "Location/Path",
        "Date Collected",
        "Collected By",
        "Verification Status"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Evidence rows (100 rows for comprehensive documentation)
    row += 1
    for i in range(1, 101):
        # Evidence ID (auto-generate)
        ws[f"A{row}"] = f"EVD-{i:03d}"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # Evidence Type dropdown
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['evidence_type'].add(ws[f"B{row}"])
        
        # Description
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Related Sheet/Row
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Location/Path
        ws[f"E{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Date Collected
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Collected By
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Verification Status dropdown
        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['verification_status'].add(ws[f"H{row}"])
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 18

    ws.freeze_panes = "A5"

# ============================================================================
# SECTION 12: APPROVAL_SIGN_OFF SHEET
# ============================================================================

def create_approval_signoff(ws, styles):
    """
    Create Approval_Sign_Off sheet.
    Formal approval workflow with 3-level sign-off.
    """
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL & SIGN-OFF"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    # ==================== ASSESSMENT SUMMARY ====================
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    summary_items = [
        ("Assessment Document:", "ISMS-IMP-A.8.23.1 - Filtering Infrastructure"),
        ("Assessment Period:", "[USER INPUT]"),
        ("Total Solutions Assessed:", "[Pull from Technology_Comparison]"),
        ("Overall Compliance Rate:", "[Pull from Capability_Requirements]"),
        ("Critical Gaps:", "[Pull from Gap_Analysis]"),
        ("Assessment Status:", "[Dropdown]"),
    ]

    for label, value in summary_items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        
        if "USER INPUT" in value:
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        elif "Dropdown" in value:
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            validations['assessment_status'].add(ws[f"B{row}"])
        
        row += 1

    # ==================== ASSESSMENT COMPLETED BY ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    completion_fields = [
        "Name:",
        "Role/Title:",
        "Department:",
        "Email:",
        "Date:",
        "Signature:",
    ]

    for field in completion_fields:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")
        row += 1

    # ==================== REVIEWED BY (ISO) ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (INFORMATION SECURITY OFFICER)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    review_fields = [
        "Name:",
        "Date:",
        "Review Notes:",
        "Recommendation:",
    ]

    for field in review_fields:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        if field == "Recommendation:":
            # Dropdown for recommendation
            rec_dv = DataValidation(
                type="list",
                formula1='"Approve,Approve with Conditions,Reject,Require Rework"',
                allow_blank=False
            )
            ws.add_data_validation(rec_dv)
            rec_dv.add(ws[f"B{row}"])
        elif field == "Review Notes:":
            # Larger cell for notes
            ws.merge_cells(f"B{row}:E{row+2}")
            ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            row += 2
        
        row += 1

    # ==================== APPROVED BY (CISO) ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    approval_fields = [
        "Name:",
        "Date:",
        "Approval Decision:",
        "Conditions/Notes:",
    ]

    for field in approval_fields:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        if field == "Approval Decision:":
            validations['approval_decision'].add(ws[f"B{row}"])
        elif field == "Conditions/Notes:":
            ws.merge_cells(f"B{row}:E{row+2}")
            ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            row += 2
        
        row += 1

    # ==================== NEXT REVIEW DETAILS ====================
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    next_review_fields = [
        "Next Review Date:",
        "Review Responsible:",
        "Special Considerations:",
    ]

    for field in next_review_fields:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        if field == "Next Review Date:":
            # Auto-calculate +3 months from today
            ws[f"B{row}"] = f'=TODAY()+90'
            ws[f"B{row}"].number_format = 'DD.MM.YYYY'
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws.freeze_panes = "A3"

# ============================================================================
# SECTION 13: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Philosophy: "En matière de grande catastrophe publique, toujours 
    privilégier la connerie au complot..." - Michel Rocard
    
    Translation: Bad ISMS implementations aren't conspiracies, they're just 
    lazy cargo cult behavior. This script chooses intelligence and organization.
    """
    print("=" * 78)
    print("ISMS-IMP-A.8.23.1 - Filtering Infrastructure Assessment Generator")
    print("ISO/IEC 27001:2022 Control A.8.23: Web Filtering")
    print("=" * 78)
    print("\n🎯 Systems Engineering Approach: Evidence-Based Compliance")
    print("📊 Vendor-Agnostic: Works with ANY filtering technology")
    print("🔒 Audit-Ready: Comprehensive evidence collection")
    print("\n" + "─" * 78)

    # Create workbook and setup styles
    print("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    print("✅ Workbook created with 10 sheets")

    # Create all sheets
    print("\n[Phase 2] Generating assessment sheets...")
    
    print("  [1/10] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    print("  ✅ Instructions complete")

    print("  [2/10] Creating Solution_Details_Template...")
    create_solution_details_template(wb["Solution_Details_Template"], styles)
    print("  ✅ Solution template complete (~70 capability assessments)")

    print("  [3/10] Creating Technology_Comparison...")
    create_technology_comparison(wb["Technology_Comparison"], styles)
    print("  ✅ Comparison matrix complete")

    print("  [4/10] Creating Capability_Requirements...")
    create_capability_requirements(wb["Capability_Requirements"], styles)
    print("  ✅ Requirements checklist complete (30 policy requirements)")

    print("  [5/10] Creating Integration_Architecture...")
    create_integration_architecture(wb["Integration_Architecture"], styles)
    print("  ✅ Integration mapping complete")

    print("  [6/10] Creating Licensing_Support...")
    create_licensing_support(wb["Licensing_Support"], styles)
    print("  ✅ License/support tracking complete")

    print("  [7/10] Creating Performance_Metrics...")
    create_performance_metrics(wb["Performance_Metrics"], styles)
    print("  ✅ Performance tracking complete")

    print("  [8/10] Creating Gap_Analysis...")
    create_gap_analysis(wb["Gap_Analysis"], styles)
    print("  ✅ Gap register complete (40 gap tracking rows)")

    print("  [9/10] Creating Evidence_Register...")
    create_evidence_register(wb["Evidence_Register"], styles)
    print("  ✅ Evidence register complete (100 evidence rows)")

    print("  [10/10] Creating Approval_Sign_Off...")
    create_approval_signoff(wb["Approval_Sign_Off"], styles)
    print("  ✅ Approval workflow complete")

    # Save workbook
    print("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.23.1_Filtering_Infrastructure_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        wb.save(filename)
        print(f"✅ SUCCESS: {filename}")
    except Exception as e:
        print(f"❌ ERROR saving workbook: {e}")
        return 1

    # Summary
    print("\n" + "=" * 78)
    print("📋 WORKBOOK STRUCTURE SUMMARY")
    print("=" * 78)
    print("\n📄 Assessment Sheets:")
    print("  • Instructions & Legend (usage guidance)")
    print("  • Solution_Details_Template (copy for each filtering solution)")
    print("  • Technology_Comparison (side-by-side comparison)")
    print("  • Capability_Requirements (policy compliance verification)")
    print("\n🔧 Operational Tracking:")
    print("  • Integration_Architecture (10 integration points)")
    print("  • Licensing_Support (license/support/update tracking)")
    print("  • Performance_Metrics (uptime/latency/incidents)")
    print("\n📊 Analysis & Governance:")
    print("  • Gap_Analysis (40 gap tracking rows)")
    print("  • Evidence_Register (100 evidence entries)")
    print("  • Approval_Sign_Off (3-level approval workflow)")
    print("\n" + "─" * 78)
    print("📈 ASSESSMENT CAPABILITIES:")
    print("  • 70+ capability assessment items per solution")
    print("  • 30 policy requirements mapped to solutions")
    print("  • 40 gap identification/remediation rows")
    print("  • 100 evidence documentation entries")
    print("  • 20 incident tracking rows")
    print("  • 12 months false positive/negative tracking")
    print("\n" + "─" * 78)
    print("🎯 KEY FEATURES:")
    print("  ✅ Vendor-agnostic (works with ANY web filtering solution)")
    print("  ✅ Comprehensive evidence collection")
    print("  ✅ Automated compliance calculations")
    print("  ✅ Gap tracking and remediation roadmap")
    print("  ✅ Multi-level approval workflow")
    print("  ✅ Quarterly review cycle support")
    print("\n" + "=" * 78)
    print("🚀 NEXT STEPS:")
    print("  1. Open the generated workbook")
    print("  2. Complete Instructions & Legend sheet first")
    print("  3. Copy Solution_Details_Template for EACH filtering solution")
    print("  4. Fill in YOUR vendor/product names (vendor-agnostic approach)")
    print("  5. Complete remaining assessment sheets")
    print("  6. Document evidence in Evidence_Register")
    print("  7. Obtain final approval via Approval_Sign_Off")
    print("\n💡 PRO TIP:")
    print("  This workbook is technology-independent. Whether you use Sophos,")
    print("  Fortigate, Zscaler, Cisco, or ANY other solution - this framework")
    print("  assesses CAPABILITIES, not brand names. That's Systems Engineering.")
    print("\n" + "=" * 78)
    print('\n"The first principle is that you must not fool yourself')
    print('— and you are the easiest person to fool." - Richard Feynman')
    print("\n🎓 This is not cargo cult ISMS. This is evidence-based compliance.")
    print("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    exit(main())