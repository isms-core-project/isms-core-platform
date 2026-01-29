#!/usr/bin/env python3
"""
ISMS-IMP-A.8.23.2 - Network Coverage Assessment Excel Generator
ISO/IEC 27001:2022 Control A.8.23 (Web Filtering - Network Coverage)

Requirements:
    sudo apt install python3-openpyxl
    
Usage:
    python3 generate_a823_2_network_coverage.py

"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches ISMS-IMP-A.8.23.2 specification
    sheets = [
        "Instructions & Legend",
        "Network_Segment_Inventory",
        "Coverage_Matrix",
        "Gap_Identification",
        "Device_Inventory",
        "Exemption_Register",
        "Coverage_Verification",
        "Evidence_Register",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    Returns style TEMPLATES (dictionaries) to avoid shared object warnings.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_protected": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_unprotected": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_planned": {
            "fill": PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        },
        "status_exempt": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        },
        "risk_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "risk_high": {
            "fill": PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),
        },
        "risk_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "risk_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell. Creates NEW objects to avoid shared warnings."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & COLUMN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """Create data validation objects for standard dropdowns."""
    validations = {
        'yes_no': DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        ),
        'yes_no_na': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'yes_no_restricted': DataValidation(
            type="list",
            formula1='"Yes,No,Restricted"',
            allow_blank=False
        ),
        'segment_type': DataValidation(
            type="list",
            formula1='"On-Premises LAN,WLAN,VPN,Cloud Endpoints,Guest,DMZ,Branch,Mobile,IoT/OT,Dev/Test,Other"',
            allow_blank=False
        ),
        'filtering_required': DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False
        ),
        'filtering_status': DataValidation(
            type="list",
            formula1='"✅ Protected,⚠️ Partial,❌ Unprotected,🔄 Planned,🚫 Exempt"',
            allow_blank=False
        ),
        'device_type': DataValidation(
            type="list",
            formula1='"Laptop,Desktop,Smartphone,Tablet,Virtual Desktop,Other"',
            allow_blank=False
        ),
        'agent_based': DataValidation(
            type="list",
            formula1='"Yes (endpoint agent),No (network-based),Hybrid"',
            allow_blank=False
        ),
        'exemption_type': DataValidation(
            type="list",
            formula1='"Full Exemption,Partial Exemption,Temporary Exemption,Category Exemption,Technical Exemption"',
            allow_blank=False
        ),
        'exemption_status': DataValidation(
            type="list",
            formula1='"Active,Expired,Revoked,Under Review"',
            allow_blank=False
        ),
        'test_method': DataValidation(
            type="list",
            formula1='"Manual Browse Test,Automated Scan,Curl/wget Test,Browser Extension,Vendor Tool,Penetration Test,Other"',
            allow_blank=False
        ),
        'test_result': DataValidation(
            type="list",
            formula1='"Pass,Fail,Partial,Inconclusive"',
            allow_blank=False
        ),
        'verification_status': DataValidation(
            type="list",
            formula1='"✅ Verified,❌ Failed,⚠️ Needs Retest"',
            allow_blank=False
        ),
        'risk_level': DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        ),
        'gap_status': DataValidation(
            type="list",
            formula1='"Open,In Progress,Resolved,Accepted Risk,Exempt"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Network Diagram,VLAN Config,Firewall Policy,VPN Config,WiFi Config,Cloud Dashboard,NAC Policy,Exemption Approval,Test Results,Monitoring Report,Change Record,Other"',
            allow_blank=False
        ),
        'evidence_verification': DataValidation(
            type="list",
            formula1='"Verified,Pending,Not Verified"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1='"Draft,Final,Requires Remediation,Re-assessment Required"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected"',
            allow_blank=False
        ),
        'recommendation': DataValidation(
            type="list",
            formula1='"Approve,Approve with Conditions,Reject,Require Rework"',
            allow_blank=False
        ),
    }

    # Add all validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)

    return validations


# ============================================================================
# SECTION 3: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet with comprehensive guidance."""
    
    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "ISMS-IMP-A.8.23.2 – Network Coverage Assessment"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "ISO/IEC 27001:2022 - Control A.8.23: Web Filtering"
    apply_style(ws["A2"], styles["subheader"])

    # Document Information Block
    row = 4
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    
    doc_info = [
        ("Document ID:", "ISMS-IMP-A.8.23.2"),
        ("Assessment Area:", "Network Coverage & Deployment Topology"),
        ("Related Policy:", "ISMS-POL-A.8.23-S2.1"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Quarterly"),
    ]
    
    row += 1
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if "USER INPUT" in value:
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    # How to Use This Workbook
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    instructions = [
        "1. Complete Network_Segment_Inventory for EVERY network segment in your environment",
        "2. Document YOUR specific network architecture (office LANs, WLANs, VPN, cloud, etc.)",
        "3. Map filtering solutions to network segments in Coverage_Matrix",
        "4. Calculate coverage percentages and identify gaps",
        "5. Document any devices/segments exempt from filtering in Exemption_Register",
        "6. Track evidence in Evidence_Register",
        "7. Review and approve via Approval_Sign_Off",
        "8. Update quarterly or after network topology changes",
        "9. Coordinate with ISMS-IMP-A.8.23.1 (Infrastructure Assessment)",
    ]

    row += 1
    for instruction in instructions:
        ws[f"A{row}"] = instruction
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1

    # Status Legend
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "STATUS LEGEND"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    legend_headers = ["Symbol", "Status", "Description", "Color"]
    for col_idx, header in enumerate(legend_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    legend_data = [
        ("✅", "Protected", "Web filtering active and operational", "Green"),
        ("⚠️", "Partial", "Partial coverage or limited protection", "Yellow"),
        ("❌", "Unprotected", "No web filtering deployed", "Red"),
        ("🔄", "Planned", "Protection planned with target date", "Blue"),
        ("🚫", "Exempt", "Approved exemption (documented)", "Gray"),
    ]

    row += 1
    for symbol, status, desc, color in legend_data:
        ws[f"A{row}"] = symbol
        ws[f"B{row}"] = status
        ws[f"C{row}"] = desc
        ws[f"D{row}"] = color
        
        if color == "Green":
            ws[f"D{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif color == "Yellow":
            ws[f"D{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif color == "Red":
            ws[f"D{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif color == "Blue":
            ws[f"D{row}"].fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        elif color == "Gray":
            ws[f"D{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        row += 1

    # Network Segment Types
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "NETWORK SEGMENT TYPES (Examples - Map YOUR environment)"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    segment_types = [
        "• On-Premises LANs (corporate office networks)",
        "• Wireless Networks (WLAN) - corporate WiFi, guest WiFi",
        "• Remote Access / VPN (remote workers, site-to-site)",
        "• Cloud-Based Endpoints (VDI, DaaS, Windows 365, virtual desktops)",
        "• Guest/Partner Networks (visitor access, vendor access)",
        "• DMZ / Extranet (public-facing services, partner connections)",
        "• Branch Offices (remote sites, field offices)",
        "• Mobile Devices (smartphones, tablets, laptops outside VPN)",
        "• IoT/OT Networks (building systems, manufacturing)",
        "• Development/Test Networks (staging, labs)",
    ]

    row += 1
    for seg_type in segment_types:
        ws[f"A{row}"] = seg_type
        row += 1

    # Acceptable Evidence
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (Examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    evidence_examples = [
        "✓ Network topology diagrams",
        "✓ VLAN configurations and documentation",
        "✓ Firewall/routing policies showing filtering enforcement",
        "✓ DHCP configurations (DNS redirection, proxy settings)",
        "✓ VPN concentrator configurations",
        "✓ Cloud service dashboards (endpoint protection)",
        "✓ WiFi controller configurations",
        "✓ Network access control (NAC) policies",
        "✓ Exemption request approvals",
        "✓ Coverage verification tests/reports",
        "✓ Network monitoring dashboards",
        "✓ Change management records for network modifications",
    ]

    row += 1
    for evidence in evidence_examples:
        ws[f"A{row}"] = evidence
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 4: NETWORK_SEGMENT_INVENTORY SHEET
# ============================================================================

def create_network_segment_inventory(ws, styles):
    """Create Network_Segment_Inventory sheet - foundation of coverage analysis."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:O1")
    ws["A1"] = "NETWORK SEGMENT INVENTORY"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:O2")
    ws["A2"] = "Document all network segments requiring web filtering protection"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        ("A", "Segment ID", 15),
        ("B", "Segment Name", 30),
        ("C", "Segment Type", 25),
        ("D", "Location/Site", 25),
        ("E", "Subnet/VLAN", 20),
        ("F", "User Count", 12),
        ("G", "Device Count", 12),
        ("H", "Internet Access?", 15),
        ("I", "Filtering Required?", 15),
        ("J", "Filtering Status", 18),
        ("K", "Filtering Solution(s)", 30),
        ("L", "Coverage %", 12),
        ("M", "Bypass Methods", 25),
        ("N", "Exemption ID", 15),
        ("O", "Evidence", 35),
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header_text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width

    # Data rows (50 segments)
    row += 1
    for i in range(1, 51):
        # Segment ID (auto-generate)
        ws[f"A{row}"] = f"SEG-{i:03d}"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # All other fields - yellow input cells
        for col in ["B", "C", "D", "E", "F", "G", "K", "L", "M", "N", "O"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Segment Type dropdown
        validations['segment_type'].add(ws[f"C{row}"])
        
        # Internet Access dropdown
        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['yes_no_restricted'].add(ws[f"H{row}"])
        
        # Filtering Required dropdown
        ws[f"I{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['filtering_required'].add(ws[f"I{row}"])
        
        # Filtering Status dropdown
        ws[f"J{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['filtering_status'].add(ws[f"J{row}"])
        
        row += 1

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 5: COVERAGE_MATRIX SHEET
# ============================================================================

def create_coverage_matrix(ws, styles):
    """Create Coverage_Matrix sheet - cross-reference segments with solutions."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "NETWORK COVERAGE MATRIX"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "Map filtering solutions to network segments"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = ["Network Segment", "Solution 1", "Solution 2", "Solution 3", "Solution 4", "Total Coverage", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows (50 segments matching inventory)
    row += 1
    for i in range(1, 51):
        # Segment reference
        ws[f"A{row}"] = f"SEG-{i:03d} [Segment Name]"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # Solution columns (B-E) - customer fills in coverage %
        for col in ["B", "C", "D", "E"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")
        
        # Total Coverage (formula - sum of solution percentages)
        ws[f"F{row}"] = f'=SUM(B{row}:E{row})'
        ws[f"F{row}"].number_format = '0"%"'
        ws[f"F{row}"].font = Font(bold=True, color="0000FF")
        
        # Status (conditional based on total coverage)
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['filtering_status'].add(ws[f"G{row}"])
        
        row += 1

    # Summary section
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "COVERAGE SUMMARY METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    summary_items = [
        ("Total segments assessed:", f'=COUNTA(A5:A54)'),
        ("Fully protected (100%):", f'=COUNTIF(F5:F54,">=100")'),
        ("Partially protected (<100%):", f'=COUNTIFS(F5:F54,">0",F5:F54,"<100")'),
        ("Unprotected (0%):", f'=COUNTIF(F5:F54,0)'),
        ("Overall Coverage Score:", f'=AVERAGE(F5:F54)&"%"'),
    ]

    for label, formula in summary_items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, size=11, color="0000FF")
        row += 1

    # Coverage Heatmap
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "COVERAGE HEATMAP"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Coverage Level"
    ws[f"B{row}"] = "Segment Count"
    ws[f"C{row}"] = "% of Total"
    ws[f"D{row}"] = "Status"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row += 1
    heatmap_data = [
        ("100% Coverage", f'=COUNTIF(F5:F54,100)', "✅ Good", "C6EFCE"),
        ("75-99% Coverage", f'=COUNTIFS(F5:F54,">=75",F5:F54,"<100")', "⚠️ Needs Attention", "FFEB9C"),
        ("50-74% Coverage", f'=COUNTIFS(F5:F54,">=50",F5:F54,"<75")', "⚠️ Risk", "FFEB9C"),
        ("<50% Coverage", f'=COUNTIFS(F5:F54,">0",F5:F54,"<50")', "❌ Critical", "FFC7CE"),
        ("0% Coverage", f'=COUNTIF(F5:F54,0)', "❌ Urgent", "FFC7CE"),
    ]

    for level, formula, status, color in heatmap_data:
        ws[f"A{row}"] = level
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        ws[f"C{row}"] = f'=IF(B{row}=0,0,B{row}/SUM(B5:B9)*100)&"%"'
        ws[f"C{row}"].font = Font(bold=True, color="0000FF")
        ws[f"D{row}"] = status
        ws[f"D{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 6: GAP_IDENTIFICATION SHEET
# ============================================================================

def create_gap_identification(ws, styles):
    """Create Gap_Identification sheet - track coverage gaps and remediation."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:K1")
    ws["A1"] = "COVERAGE GAPS & REMEDIATION TRACKING"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = "Identify unprotected/partially protected segments and plan remediation"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        "Gap ID", "Segment ID", "Segment Name", "Current Coverage", "Gap Description",
        "Risk Level", "Impact", "Remediation Plan", "Owner", "Target Date", "Status"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Gap tracking rows (30 rows)
    row += 1
    gap_start_row = row
    
    for i in range(1, 31):
        # Gap ID
        ws[f"A{row}"] = f"GAP-{i:03d}"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # All other fields
        for col in ["B", "C", "D", "E", "G", "H", "I", "J"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Risk Level dropdown
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['risk_level'].add(ws[f"F{row}"])
        
        # Status dropdown
        ws[f"K{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['gap_status'].add(ws[f"K{row}"])
        
        row += 1
    
    gap_end_row = row - 1

    # Gap Summary Metrics
    row += 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "GAP SUMMARY METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Risk Level"
    ws[f"B{row}"] = "Gap Count"
    ws[f"C{row}"] = "% of Total"
    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row += 1
    risk_levels = [
        ("Critical", "C00000", "FFFFFF"),
        ("High", "FF6666", "000000"),
        ("Medium", "FFEB9C", "000000"),
        ("Low", "C6EFCE", "000000"),
    ]

    for risk, bg_color, fg_color in risk_levels:
        ws[f"A{row}"] = risk
        ws[f"A{row}"].font = Font(bold=True, color=fg_color)
        ws[f"A{row}"].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        ws[f"B{row}"] = f'=COUNTIF(F{gap_start_row}:F{gap_end_row},"{risk}")'
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        
        ws[f"C{row}"] = f'=IF(B{row}=0,0,B{row}/COUNTA(F{gap_start_row}:F{gap_end_row})*100)&"%"'
        ws[f"C{row}"].font = Font(bold=True, color="0000FF")
        
        row += 1

    # Overall summary
    row += 1
    overall_items = [
        ("Total gaps identified:", f'=COUNTA(B{gap_start_row}:B{gap_end_row})'),
        ("Gaps resolved:", f'=COUNTIF(K{gap_start_row}:K{gap_end_row},"Resolved")'),
        ("Gaps remaining:", f'=COUNTA(B{gap_start_row}:B{gap_end_row})-COUNTIF(K{gap_start_row}:K{gap_end_row},"Resolved")'),
    ]

    for label, formula in overall_items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, size=11, color="C00000" if "remaining" in label else "0000FF")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 15

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 7: DEVICE_INVENTORY SHEET
# ============================================================================

def create_device_inventory(ws, styles):
    """Create Device_Inventory sheet - endpoint-level tracking."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:J1")
    ws["A1"] = "DEVICE-LEVEL FILTERING INVENTORY"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Track endpoint protection for mobile/remote devices"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        ("A", "Device ID", 12),
        ("B", "Device Type", 15),
        ("C", "OS", 15),
        ("D", "User/Owner", 25),
        ("E", "Primary Network", 25),
        ("F", "Filtering Solution", 25),
        ("G", "Agent-Based?", 18),
        ("H", "Status", 15),
        ("I", "Last Verified", 15),
        ("J", "Evidence", 35),
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header_text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width

    # Device rows (100 rows)
    row += 1
    for i in range(1, 101):
        # Device ID
        ws[f"A{row}"] = f"DEV-{i:03d}"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # Device Type dropdown
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['device_type'].add(ws[f"B{row}"])
        
        # All other text fields
        for col in ["C", "D", "E", "F", "I", "J"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Agent-Based dropdown
        ws[f"G{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['agent_based'].add(ws[f"G{row}"])
        
        # Status dropdown
        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['filtering_status'].add(ws[f"H{row}"])
        
        row += 1

    # Device Summary
    row += 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "DEVICE SUMMARY BY TYPE"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Device Type"
    ws[f"B{row}"] = "Total"
    ws[f"C{row}"] = "Protected"
    ws[f"D{row}"] = "Unprotected"
    ws[f"E{row}"] = "Coverage %"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row += 1
    device_types = ["Laptop", "Desktop", "Smartphone", "Tablet", "Virtual Desktop", "Other"]
    
    for dev_type in device_types:
        ws[f"A{row}"] = dev_type
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = f'=COUNTIF(B5:B104,"{dev_type}")'
        ws[f"C{row}"] = f'=COUNTIFS(B5:B104,"{dev_type}",H5:H104,"✅ Protected")'
        ws[f"D{row}"] = f'=COUNTIFS(B5:B104,"{dev_type}",H5:H104,"❌ Unprotected")'
        ws[f"E{row}"] = f'=IF(B{row}=0,0,C{row}/B{row}*100)&"%"'
        for col in ["B", "C", "D", "E"]:
            ws[f"{col}{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 8: EXEMPTION_REGISTER SHEET
# ============================================================================

def create_exemption_register(ws, styles):
    """Create Exemption_Register sheet - approved exceptions documentation."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:K1")
    ws["A1"] = "FILTERING EXEMPTION REGISTER"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = "Document approved exceptions to web filtering requirements"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        "Exemption ID", "Segment/Device", "Exemption Type", "Business Justification",
        "Risk Assessment", "Compensating Controls", "Approved By", "Approval Date",
        "Review Date", "Status", "Evidence"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Exemption rows (20 rows)
    row += 1
    for i in range(1, 21):
        # Exemption ID
        ws[f"A{row}"] = f"EXM-{i:03d}"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # Segment/Device
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Exemption Type dropdown
        ws[f"C{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['exemption_type'].add(ws[f"C{row}"])
        
        # Text fields
        for col in ["D", "E", "F", "G", "K"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Dates
        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"I{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Status dropdown
        ws[f"J{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['exemption_status'].add(ws[f"J{row}"])
        
        row += 1

    # Exemption Summary
    row += 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "EXEMPTION SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f"A{row}"] = "Exemption Type"
    ws[f"B{row}"] = "Active Count"
    ws[f"C{row}"] = "Expired/Overdue"
    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row += 1
    exemption_types = [
        "Full Exemption",
        "Partial Exemption",
        "Temporary Exemption",
        "Category Exemption",
        "Technical Exemption",
    ]

    for exm_type in exemption_types:
        ws[f"A{row}"] = exm_type
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = f'=COUNTIFS(C5:C24,"{exm_type}",J5:J24,"Active")'
        ws[f"C{row}"] = f'=COUNTIFS(C5:C24,"{exm_type}",J5:J24,"Expired")'
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        ws[f"C{row}"].font = Font(bold=True, color="C00000")
        row += 1

    # Alert section
    row += 2
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "⚠️ ALERT: Exemptions >12 months old require re-approval | >10% of segments exempt = escalate to CISO"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="C00000")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 30

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 9: COVERAGE_VERIFICATION SHEET
# ============================================================================

def create_coverage_verification(ws, styles):
    """Create Coverage_Verification sheet - testing documentation."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:J1")
    ws["A1"] = "COVERAGE VERIFICATION LOG"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Test and verify filtering effectiveness per network segment"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        ("A", "Verification ID", 15),
        ("B", "Segment ID", 12),
        ("C", "Segment Name", 25),
        ("D", "Test Date", 15),
        ("E", "Tester", 20),
        ("F", "Test Method", 20),
        ("G", "Test Results", 30),
        ("H", "Issues Found", 35),
        ("I", "Status", 15),
        ("J", "Next Test Date", 15),
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header_text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width

    # Verification rows (50 rows)
    row += 1
    for i in range(1, 51):
        # Verification ID
        ws[f"A{row}"] = f"VER-{i:03d}"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # Segment ID, Segment Name, Tester, Test Results, Issues
        for col in ["B", "C", "E", "G", "H"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Test Date
        ws[f"D{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Test Method dropdown
        ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['test_method'].add(ws[f"F{row}"])
        
        # Status dropdown
        ws[f"I{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['verification_status'].add(ws[f"I{row}"])
        
        # Next Test Date (auto-calculate +90 days)
        ws[f"J{row}"] = f'=IF(D{row}="","",D{row}+90)'
        ws[f"J{row}"].number_format = 'DD.MM.YYYY'
        
        row += 1

    # Test Checklist Template
    row += 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "TEST CHECKLIST (Verify for each segment)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    checklist_items = [
        "☐ Malicious URL blocked (test with EICAR or known bad domains)",
        "☐ Phishing site blocked (test with PhishTank samples)",
        "☐ Known malware download blocked",
        "☐ HTTPS sites properly filtered (if HTTPS inspection enabled)",
        "☐ Logging working (blocked attempts logged)",
        "☐ Bypass methods ineffective (proxy bypass, VPN bypass, etc.)",
        "☐ User experience acceptable (latency, false positives)",
        "☐ Exemptions working as intended",
    ]

    row += 1
    for item in checklist_items:
        ws[f"A{row}"] = item
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 20
        row += 1

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 10: EVIDENCE_REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence_Register sheet - 100 row evidence repository."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence supporting network coverage assessment"
    apply_style(ws["A2"], styles["subheader"])

    # Column headers
    row = 4
    headers = [
        ("A", "Evidence ID", 12),
        ("B", "Evidence Type", 20),
        ("C", "Description", 40),
        ("D", "Related Sheet/Row", 20),
        ("E", "Location/Path", 50),
        ("F", "Date Collected", 15),
        ("G", "Collected By", 20),
        ("H", "Verification Status", 18),
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}{row}"]
        cell.value = header_text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width

    # Evidence rows (100 rows)
    row += 1
    for i in range(1, 101):
        # Evidence ID
        ws[f"A{row}"] = f"EVD-{i:03d}"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        
        # Evidence Type dropdown
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['evidence_type'].add(ws[f"B{row}"])
        
        # Text fields
        for col in ["C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Verification Status dropdown
        ws[f"H{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        validations['evidence_verification'].add(ws[f"H{row}"])
        
        row += 1

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 11: APPROVAL_SIGN_OFF SHEET
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create Approval_Sign_Off sheet with 3-level approval workflow."""
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL & SIGN-OFF"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    # Assessment Summary
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    summary_items = [
        ("Assessment Document:", "ISMS-IMP-A.8.23.2 - Network Coverage"),
        ("Assessment Period:", "[USER INPUT]"),
        ("Total Segments Assessed:", "[Formula from Network_Segment_Inventory]"),
        ("Fully Protected:", "[Formula - 100% coverage]"),
        ("Partially Protected:", "[Formula - <100% coverage]"),
        ("Unprotected:", "[Formula - 0% coverage]"),
        ("Exempted:", "[Formula from Exemption_Register]"),
        ("Overall Coverage Score:", "[Formula]%"),
        ("Critical Gaps:", "[Formula from Gap_Identification]"),
        ("Assessment Status:", "[Dropdown]"),
    ]

    for label, value in summary_items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        
        if "USER INPUT" in value:
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        elif "Dropdown" in value:
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            validations['assessment_status'].add(ws[f"B{row}"])
        
        row += 1

    # Assessment Completed By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    completion_fields = ["Name:", "Role/Title:", "Department:", "Email:", "Date:", "Signature:"]
    for field in completion_fields:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    # Reviewed By - Network Administrator
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (NETWORK ADMINISTRATOR)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name:", "Date:", "Review Notes:"]:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        if field == "Review Notes:":
            ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(f"B{row}:E{row+2}")
            row += 2
        row += 1

    ws[f"A{row}"] = "Recommendation:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    validations['recommendation'].add(ws[f"B{row}"])
    row += 1

    # Reviewed By - Information Security Officer
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (INFORMATION SECURITY OFFICER)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name:", "Date:", "Review Notes:"]:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        if field == "Review Notes:":
            ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(f"B{row}:E{row+2}")
            row += 2
        row += 1

    ws[f"A{row}"] = "Recommendation:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    validations['recommendation'].add(ws[f"B{row}"])
    row += 1

    # Approved By - CISO
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name:", "Date:", "Approval Decision:"]:
        ws[f"A{row}"] = field
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        if field == "Approval Decision:":
            validations['approval_decision'].add(ws[f"B{row}"])
        row += 1

    ws[f"A{row}"] = "Conditions/Notes:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row+2}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    row += 3

    # Next Review Details
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    next_review_fields = [
        ("Next Review Date:", "=TODAY()+90"),
        ("Review Responsible:", "[USER INPUT]"),
        ("Special Considerations:", "[USER INPUT]"),
    ]

    for label, value in next_review_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        if "USER INPUT" in value:
            pass  # Already yellow
        elif "TODAY" in value:
            ws[f"B{row}"] = value
            ws[f"B{row}"].number_format = 'DD.MM.YYYY'
        else:
            ws[f"B{row}"] = value
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    Network coverage is where the rubber meets the road. You can have the best
    filtering technology in the world, but if it only covers 60% of your network,
    you have a 40% gap for attackers to exploit.
    
    This assessment answers: WHERE is filtering applied, and WHERE are the gaps?
    """
    print("=" * 78)
    print("ISMS-IMP-A.8.23.2 - Network Coverage Assessment Generator")
    print("ISO/IEC 27001:2022 Control A.8.23: Web Filtering")
    print("=" * 78)
    print("\n🌐 Network Coverage: Verify filtering across ALL network segments")
    print("📊 Topology-Agnostic: Works with ANY network architecture")
    print("🔍 Gap Identification: Find unprotected segments")
    print("\n" + "─" * 78)

    # Create workbook and setup styles
    print("\n[Phase 1] Initializing workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    print("✅ Workbook created with 9 sheets")

    # Create all sheets
    print("\n[Phase 2] Generating assessment sheets...")
    
    print("  [1/9] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    print("  ✅ Instructions complete")

    print("  [2/9] Creating Network_Segment_Inventory...")
    create_network_segment_inventory(wb["Network_Segment_Inventory"], styles)
    print("  ✅ Segment inventory complete (50 segments)")

    print("  [3/9] Creating Coverage_Matrix...")
    create_coverage_matrix(wb["Coverage_Matrix"], styles)
    print("  ✅ Coverage matrix complete (cross-reference map)")

    print("  [4/9] Creating Gap_Identification...")
    create_gap_identification(wb["Gap_Identification"], styles)
    print("  ✅ Gap tracking complete (30 gap rows)")

    print("  [5/9] Creating Device_Inventory...")
    create_device_inventory(wb["Device_Inventory"], styles)
    print("  ✅ Device inventory complete (100 devices)")

    print("  [6/9] Creating Exemption_Register...")
    create_exemption_register(wb["Exemption_Register"], styles)
    print("  ✅ Exemption register complete (20 exemptions)")

    print("  [7/9] Creating Coverage_Verification...")
    create_coverage_verification(wb["Coverage_Verification"], styles)
    print("  ✅ Verification log complete (50 tests)")

    print("  [8/9] Creating Evidence_Register...")
    create_evidence_register(wb["Evidence_Register"], styles)
    print("  ✅ Evidence register complete (100 evidence rows)")

    print("  [9/9] Creating Approval_Sign_Off...")
    create_approval_signoff(wb["Approval_Sign_Off"], styles)
    print("  ✅ Approval workflow complete")

    # Save workbook
    print("\n[Phase 3] Finalizing and saving workbook...")
    filename = f"ISMS-IMP-A.8.23.2_Network_Coverage_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        wb.save(filename)
        print(f"✅ SUCCESS: {filename}")
    except Exception as e:
        print(f"❌ ERROR saving workbook: {e}")
        return 1

    # Summary
    print("\n" + "=" * 78)
    print("📋 WORKBOOK STRUCTURE SUMMARY")
    print("=" * 78)
    print("\n📄 Core Assessment:")
    print("  • Instructions & Legend (usage guidance + segment types)")
    print("  • Network_Segment_Inventory (50 segments - foundation)")
    print("  • Coverage_Matrix (cross-reference solutions to segments)")
    print("  • Gap_Identification (30 gap tracking rows)")
    print("\n🔍 Supplementary Tracking:")
    print("  • Device_Inventory (100 endpoint devices)")
    print("  • Exemption_Register (20 approved exemptions)")
    print("  • Coverage_Verification (50 test entries)")
    print("\n📊 Governance:")
    print("  • Evidence_Register (100 evidence entries)")
    print("  • Approval_Sign_Off (3-level approval workflow)")
    print("\n" + "─" * 78)
    print("📈 ASSESSMENT CAPABILITIES:")
    print("  • 50 network segments documented")
    print("  • Cross-reference with 4 filtering solutions")
    print("  • 100 endpoint devices tracked")
    print("  • 30 gap identification/remediation rows")
    print("  • 20 exemption approvals")
    print("  • 50 verification test entries")
    print("  • Auto-calculated coverage percentages")
    print("  • Coverage heatmap analysis")
    print("\n" + "─" * 78)
    print("🎯 KEY FEATURES:")
    print("  ✅ Topology-agnostic (map ANY network architecture)")
    print("  ✅ Comprehensive segment coverage tracking")
    print("  ✅ Automated coverage percentage calculations")
    print("  ✅ Gap prioritization by risk level")
    print("  ✅ Exemption management with approval workflow")
    print("  ✅ Verification testing documentation")
    print("  ✅ Multi-level approval process")
    print("\n" + "=" * 78)
    print("🚀 NEXT STEPS:")
    print("  1. Open the generated workbook")
    print("  2. Document ALL network segments in Network_Segment_Inventory")
    print("  3. Map solutions to segments in Coverage_Matrix")
    print("  4. Calculate coverage percentages (auto-calculated)")
    print("  5. Identify gaps in Gap_Identification")
    print("  6. Document exemptions if needed")
    print("  7. Verify coverage with testing (Coverage_Verification)")
    print("  8. Obtain approvals via Approval_Sign_Off")
    print("\n💡 PRO TIP:")
    print("  Coverage theater says 'we have web filtering.'")
    print("  Systems engineering says 'we have 97.3% coverage across 47 segments")
    print("  with 3 approved exemptions and quarterly verification testing.'")
    print("\n📐 COVERAGE FORMULA:")
    print("  Network Coverage Score = Σ(Protected Segments) / Total Segments * 100%")
    print("\n" + "=" * 78)
    print('\n"Cargo cult ISMS: We have filtering. Somewhere. Maybe."')
    print('"Systems Engineering: Verified 97.3% coverage with evidence."')
    print("\n🎓 Evidence > Theater")
    print("=" * 78 + "\n")

    return 0


if __name__ == "__main__":
    exit(main())