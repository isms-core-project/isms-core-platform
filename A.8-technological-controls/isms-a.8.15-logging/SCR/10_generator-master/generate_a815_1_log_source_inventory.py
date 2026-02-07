#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.15.1 - Log Source Inventory Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.15: Logging
Assessment Domain 1 of 4: System Inventory and Log Source Cataloging

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific logging infrastructure, system architecture, and
assessment requirements.

Key customization areas:
1. System types and categories (match your infrastructure taxonomy)
2. Log source priorities and criticality levels (adapt to your risk profile)
3. Regulatory scope mappings (specific to your compliance requirements)
4. Data classification schemes (based on your information security policy)
5. Compliance scoring criteria (aligned with your ISMS maturity targets)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.15 Logging Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for documenting
and evaluating log sources across the organization's entire IT infrastructure.

**Purpose:**
Enables systematic inventory and assessment of all systems requiring logging
against ISO 27001:2022 Control A.8.15 requirements, supporting evidence-based
validation of logging coverage and completeness.

**Assessment Scope:**
- System inventory (servers, workstations, network devices, applications)
- Log event types by system (authentication, authorization, administrative)
- Authentication logging capabilities and implementation
- Authorization and access control logging
- Administrative activity logging
- Security event logging (intrusions, malware, anomalies)
- Application and database logging
- Network device logging (firewalls, routers, switches, IDS/IPS)
- Gap analysis for systems without adequate logging
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and logging standards
2. System Inventory - Complete catalog of in-scope systems
3. Log Event Types by System - Event type mapping per system
4. Authentication Logging - Login/logout event assessment
5. Authorization & Access - Permission and access control logging
6. Administrative Activity - Privileged operation logging
7. Security Event Logging - Security-relevant event assessment
8. Application & Database - Application-specific logging
9. Network Device Logging - Network infrastructure logging
10. Gap Analysis - Systems lacking adequate logging
11. Evidence Register - Audit evidence tracking and documentation
12. Summary Dashboard - Compliance metrics and KPIs
13. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with dropdown lists for consistency
- Conditional formatting for compliance status visualization
- Automated gap identification for systems without logging
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with SIEM and log management platforms

**Integration:**
This assessment feeds into the A.8.15.5 Compliance Dashboard, which
consolidates data from all four logging assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a815_1_log_source_inventory.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a815_1_log_source_inventory.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a815_1_log_source_inventory.py --date 20250124

Output:
    File: ISMS_A_8_15_1_Log_Source_Inventory_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize system categories to match your environment
    2. Inventory all systems in scope for logging requirements
    3. Complete log source assessments for each system
    4. Validate logging capabilities against requirements
    5. Review gap analysis for systems without adequate logging
    6. Define remediation actions with timelines
    7. Collect and link audit evidence (configurations, screenshots)
    8. Obtain stakeholder approvals
    9. Feed results into A.8.15.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.15
Assessment Domain:    1 of 4 (System Inventory and Log Source Cataloging)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.15: Logging Policy (Governance)
    - ISMS-IMP-A.8.15.1: Log Source Inventory Implementation Guide
    - ISMS-IMP-A.8.15.2: Log Collection & Centralization Assessment (Domain 2)
    - ISMS-IMP-A.8.15.3: Log Protection & Retention Assessment (Domain 3)
    - ISMS-IMP-A.8.15.4: Log Analysis & Review Assessment (Domain 4)
    - ISMS-IMP-A.8.15.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 24.01.2025
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.15.1 specification
    - Supports comprehensive log source inventory and cataloging
    - Integrated with A.8.15.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Logging Standards:**
Logging requirements evolve with threat landscape changes. Review industry
standards (NIST SP 800-92, CIS Critical Security Controls, MITRE ATT&CK)
quarterly and update assessment criteria accordingly. Ensure logging coverage
aligns with current security best practices.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of logging capabilities for all critical
and high-risk systems.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- System inventory and network topology
- Security control implementation details
- Vulnerability information and security gaps
- Regulatory compliance scope

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Check for new systems added to infrastructure
- Semi-annually: Update logging requirements based on threat landscape
- Annually: Complete reassessment of all systems
- Ad-hoc: When infrastructure changes or new compliance requirements emerge

**Quality Assurance:**
Have logging/SIEM SMEs and security operations engineers validate assessments
before using results for compliance reporting or remediation decisions.

**Regulatory Alignment:**
Ensure logging requirements align with applicable regulatory requirements:
- Payment processing: PCI DSS logging and monitoring requirements
- Healthcare: HIPAA audit trail requirements
- Finance: Regional financial services logging requirements
- Government: Jurisdiction-specific audit logging mandates

Customize assessment criteria to include regulatory-specific requirements.

**SIEM Integration:**
This assessment identifies log sources that should be forwarded to SIEM.
Coordinate with SOC team to ensure identified log sources are configured
for centralized collection and analysis.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •

# Document identification constants
DOCUMENT_ID = "ISMS-IMP-A.8.15.1"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_Log_Source_Inventory_{GENERATED_TIMESTAMP}.xlsx"
CONTROL_REF = "ISO/IEC 27001:2022 - Control A.8.15: Logging"

from openpyxl.chart import PieChart, BarChart, Reference


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches ISMS-IMP-A.8.15.1 specification
    sheets = [
        "Instructions & Legend",
        "System Inventory",
        "Log Event Types by System",
        "Authentication Logging",
        "Authorization & Access",
        "Administrative Activity",
        "Security Event Logging",
        "Application & Database",
        "Network Device Logging",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval & Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    Returns style TEMPLATES (dictionaries) to avoid shared object warnings.
    Each cell gets its own style instance to prevent Excel repair issues.
    
    Philosophy: "The first principle is that you must not fool yourself"
    - Feynman. Don't cargo cult shared styles - create fresh ones!
    """
    return {
        'header_main': {
            'font': {'name': 'Calibri', 'size': 14, 'bold': True, 'color': 'FFFFFF'},
            'fill': {'start_color': '003366', 'end_color': '003366', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'header_sub': {
            'font': {'name': 'Calibri', 'size': 11, 'bold': True, 'color': 'FFFFFF'},
            'fill': {'start_color': '4472C4', 'end_color': '4472C4', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'column_header': {
            'font': {'name': 'Calibri', 'size': 10, 'bold': True, 'color': '000000'},
            'fill': {'start_color': 'D9D9D9', 'end_color': 'D9D9D9', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'input_cell': {
            'fill': {'start_color': 'FFFFCC', 'end_color': 'FFFFCC', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'left', 'vertical': 'top', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'example_cell': {
            'fill': {'start_color': 'E7E6E6', 'end_color': 'E7E6E6', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'italic': True, 'color': '666666'},
            'alignment': {'horizontal': 'left', 'vertical': 'top'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'formula_cell': {
            'fill': {'start_color': 'E7E6E6', 'end_color': 'E7E6E6', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'status_compliant': {
            'fill': {'start_color': 'C6EFCE', 'end_color': 'C6EFCE', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'bold': True, 'color': '006100'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'status_partial': {
            'fill': {'start_color': 'FFEB9C', 'end_color': 'FFEB9C', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'bold': True, 'color': '9C6500'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'status_noncompliant': {
            'fill': {'start_color': 'FFC7CE', 'end_color': 'FFC7CE', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'bold': True, 'color': '9C0006'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'info_cell': {
            'fill': {'start_color': 'E7E6E6', 'end_color': 'E7E6E6', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'color': '000000'},
            'alignment': {'horizontal': 'left', 'vertical': 'top', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        }
    }


def apply_style(cell, style_template):
    """
    Apply a style template to a cell.
    Creates NEW objects for each cell to avoid Excel repair warnings.
    """
    if 'font' in style_template:
        cell.font = Font(**style_template['font'])
    if 'fill' in style_template:
        cell.fill = PatternFill(**style_template['fill'])
    if 'alignment' in style_template:
        cell.alignment = Alignment(**style_template['alignment'])
    if 'border' in style_template:
        cell.border = Border(
            left=Side(style=style_template['border'].get('left', 'thin')),
            right=Side(style=style_template['border'].get('right', 'thin')),
            top=Side(style=style_template['border'].get('top', 'thin')),
            bottom=Side(style=style_template['border'].get('bottom', 'thin'))
        )


def set_column_widths(ws, widths):
    """Set column widths from a dictionary {column_letter: width}."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

# ============================================================================
# SECTION 2: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """
    Create Instructions & Legend sheet with document info and usage guide.
    
    "If you can't explain it simply, you don't understand it well enough"
    - Einstein. So let's make these instructions crystal clear!
    """
    
    # Main header
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{DOCUMENT_ID}  -  Log Source Inventory Assessment\n{CONTROL_REF}"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    # Document Information Block
    row = 4
    info_fields = [
        ("Document ID:", "ISMS-IMP-A.8.15.1"),
        ("Assessment Area:", "Log Source Inventory and Completeness"),
        ("Related Policy:", "ISMS-POL-A.8.15-S2.1 (Event Logging Requirements)"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT - DD.MM.YYYY]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Annual (full assessment), Quarterly (updates)"),
    ]
    
    for label, value in info_fields:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = value
        
        if "[USER INPUT" in value:
            apply_style(ws[f'B{row}'], styles['input_cell'])
            if "Date" in label:
                ws[f'B{row}'].number_format = 'DD.MM.YYYY'
        else:
            apply_style(ws[f'B{row}'], styles['info_cell'])
        
        row += 1
    
    # How to Use This Workbook
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "HOW TO USE THIS WORKBOOK"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 1
    instructions = [
        "1. Complete the 'System Inventory' sheet for all in-scope systems",
        "2. For each system, document log sources and event types logged",
        "3. Use dropdowns to ensure consistent data entry",
        "4. Yellow cells require user input, gray cells are auto-calculated",
        "5. Complete compliance checklist for each log source type",
        "6. Review Summary Dashboard for overall compliance status",
        "7. Document gaps in Gap Analysis sheet",
        "8. Develop remediation plan for identified gaps",
        "9. Obtain approvals in Approval & Sign-Off sheet"
    ]
    
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        row += 1
    
    # Legend
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "LEGEND - CELL COLOR CODING"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 1
    legend_items = [
        ("Yellow (FFFFCC)", "User input required"),
        ("Green (C6EFCE)", "Compliant / Complete"),
        ("Red (FFC7CE)", "Non-compliant / Missing"),
        ("Gray (E7E6E6)", "Auto-calculated / Reference data"),
        ("Blue (4472C4)", "Instructions / Headers"),
    ]
    
    for color_desc, meaning in legend_items:
        ws[f'A{row}'] = color_desc
        ws[f'B{row}'] = meaning
        
        # Apply the actual color to column A
        if "Yellow" in color_desc:
            ws[f'A{row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        elif "Green" in color_desc:
            ws[f'A{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif "Red" in color_desc:
            ws[f'A{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif "Gray" in color_desc:
            ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        elif "Blue" in color_desc:
            ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            ws[f'A{row}'].font = Font(color='FFFFFF', bold=True)
        
        ws[f'A{row}'].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        row += 1
    
    # Key Definitions
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "KEY DEFINITIONS"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 1
    definitions = [
        ("Log Source", "Any system, application, or device generating event logs"),
        ("Event Type", "Category of logged activity (authentication, authorization, etc.)"),
        ("SIEM", "Security Information and Event Management system"),
        ("Hot Storage", "Real-time searchable log storage (0-90 days typical)"),
        ("Warm Storage", "Compressed, slower-access storage (91-365 days typical)"),
        ("Cold Storage", "Archive storage (>1 year)"),
        ("WORM", "Write-Once-Read-Many (immutable storage)"),
        ("Syslog", "Standard logging protocol (RFC 5424)"),
        ("CEF", "Common Event Format"),
        ("EVTX", "Windows Event Log format"),
    ]
    
    for term, definition in definitions:
        ws[f'A{row}'] = term
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = definition
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        row += 1
    
    # Set column widths
    set_column_widths(ws, {
        'A': 25,
        'B': 60,
        'C': 15,
        'D': 15,
        'E': 15,
        'F': 15
    })
    
    ws.freeze_panes = 'A3'

# ============================================================================
# SECTION 3: SYSTEM INVENTORY SHEET
# ============================================================================

def create_system_inventory_sheet(ws, styles):
    """
    Create System Inventory sheet - the foundation of log source catalog.
    
    "You can't improve what you don't measure, and you can't measure 
    what you don't catalog." - Adapted from Peter Drucker
    """
    
    # Header
    ws.merge_cells('A1:Q1')
    ws['A1'] = "SYSTEM INVENTORY - Log Source Catalog"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    # Instructions
    ws.merge_cells('A2:Q2')
    ws['A2'] = "Document all systems in scope for logging. Include servers, network devices, applications, cloud services, and security tools."
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Column headers (Row 7)
    headers = [
        ("A", "System ID", 15),
        ("B", "System Name", 30),
        ("C", "System Type", 20),
        ("D", "Operating System / Platform", 25),
        ("E", "Environment", 15),
        ("F", "Data Classification", 18),
        ("G", "Business Criticality", 18),
        ("H", "Regulatory Scope", 20),
        ("I", "Logging Priority", 15),
        ("J", "System Owner", 25),
        ("K", "Owner Email", 30),
        ("L", "Hostname / FQDN", 30),
        ("M", "Primary IP", 15),
        ("N", "Location", 20),
        ("O", "Logging Enabled", 15),
        ("P", "Forwarding to SIEM", 18),
        ("Q", "Compliance Status", 18),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row (Row 8 - Gray italic)
    row = 8
    example_data = [
        "SYS-001", "web-prod-01", "Server", "Linux", "Production",
        "Confidential", "Critical (T1)", "PCI DSS, GDPR", "P1-Critical",
        "John Doe", "jdoe@example.com", "web-prod-01.example.com",
        "10.0.1.50", "DC1", "Yes", "Yes", "Compliant"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-100 = 92 rows)
    for data_row in range(9, 101):
        # Column A: Auto-generate System ID
        ws[f'A{data_row}'] = f'=IF(B{data_row}<>"","SYS-"&TEXT(ROW()-8,"000"),"")'
        apply_style(ws[f'A{data_row}'], styles['formula_cell'])
        
        # Columns B-P: Input cells
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column Q: Compliance Status Formula
        ws[f'Q{data_row}'] = f'=IF(B{data_row}="","",IF(AND(O{data_row}="Yes",P{data_row}="Yes"),"Compliant",IF(OR(O{data_row}="Partial",P{data_row}="Planned"),"Partial","Non-Compliant")))'
        apply_style(ws[f'Q{data_row}'], styles['formula_cell'])
    
    # Data validations
    system_type_dv = DataValidation(type="list",
        formula1='"Server,Network Device,Security Appliance,Application,Cloud Service,Database,Other"',
        allow_blank=False)
    ws.add_data_validation(system_type_dv)
    system_type_dv.add(f'C9:C100')
    
    os_platform_dv = DataValidation(type="list",
        formula1='"Windows,Linux,Unix,Network OS,Cloud,Application Platform,Other"',
        allow_blank=False)
    ws.add_data_validation(os_platform_dv)
    os_platform_dv.add(f'D9:D100')
    
    environment_dv = DataValidation(type="list",
        formula1='"Production,Staging,Development,Test"',
        allow_blank=False)
    ws.add_data_validation(environment_dv)
    environment_dv.add(f'E9:E100')
    
    data_class_dv = DataValidation(type="list",
        formula1='"Public,Internal,Confidential,Restricted"',
        allow_blank=False)
    ws.add_data_validation(data_class_dv)
    data_class_dv.add(f'F9:F100')
    
    criticality_dv = DataValidation(type="list",
        formula1='"🔴 Critical (T1),🟡 High (T2),🟢 Medium (T3),⚫ Low (T4)"',
        allow_blank=False)
    ws.add_data_validation(criticality_dv)
    criticality_dv.add(f'G9:G100')
    
    priority_dv = DataValidation(type="list",
        formula1='"🔴 P1-Critical,🟡 P2-High,🟢 P3-Medium,⚫ P4-Low"',
        allow_blank=False)
    ws.add_data_validation(priority_dv)
    priority_dv.add(f'I9:I100')
    
    logging_enabled_dv = DataValidation(type="list",
        formula1='"\u2705 Yes,\u274C No,\u26A0\uFE0F Partial,❓ Unknown"',
        allow_blank=False)
    ws.add_data_validation(logging_enabled_dv)
    logging_enabled_dv.add(f'O9:O100')
    
    forwarding_dv = DataValidation(type="list",
        formula1='"\u2705 Yes,\u274C No,⏳ Planned,➖ N/A"',
        allow_blank=False)
    ws.add_data_validation(forwarding_dv)
    forwarding_dv.add(f'P9:P100')
    
    ws.freeze_panes = 'A8'

# ============================================================================
# SECTION 4: LOG EVENT TYPES BY SYSTEM SHEET
# ============================================================================

def create_log_event_types_sheet(ws, styles):
    """
    Create Log Event Types sheet - documents what each system logs.
    
    "The devil is in the details, but so is salvation." - Hyman Rickover
    Let's get granular about what's actually being logged!
    """
    
    # Header
    ws.merge_cells('A1:S1')
    ws['A1'] = "LOG EVENT TYPES BY SYSTEM"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    # Instructions
    ws.merge_cells('A2:S2')
    ws['A2'] = "For each system, document which event types are logged. This assesses logging completeness per ISMS-POL-A.8.15-S2.1"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Column headers (Row 7)
    headers = [
        ("A", "System ID", 15),
        ("B", "System Name", 30),
        ("C", "Authentication Events", 15),
        ("D", "Authorization Events", 15),
        ("E", "Administrative Actions", 15),
        ("F", "Security Events", 15),
        ("G", "Application Events", 15),
        ("H", "System Events", 15),
        ("I", "Network Events", 15),
        ("J", "Database Events", 15),
        ("K", "Log Format", 15),
        ("L", "Timestamp Format", 18),
        ("M", "Timezone", 15),
        ("N", "Est. Daily Volume (MB)", 20),
        ("O", "Retention Period (months)", 20),
        ("P", "Storage Tier", 15),
        ("Q", "Protection Mechanisms", 20),
        ("R", "Event Types Completeness", 20),
        ("S", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row (Row 8)
    row = 8
    example_data = [
        "SYS-001", "web-prod-01", "Yes", "Yes", "Yes", "Yes", "Yes",
        "Partial", "Partial", "N/A", "Syslog", "ISO 8601", "UTC",
        "150", "12", "Hot", "Access Controls, Encryption", "87.5%",
        "Web server logs include app events"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-100)
    for data_row in range(9, 101):
        # Column A: System ID dropdown (links to System Inventory)
        apply_style(ws[f'A{data_row}'], styles['input_cell'])
        
        # Column B: VLOOKUP System Name
        ws[f'B{data_row}'] = f'=IF(A{data_row}="","",IFERROR(VLOOKUP(A{data_row},\'System Inventory\'!A:B,2,FALSE),"Not Found"))'
        apply_style(ws[f'B{data_row}'], styles['formula_cell'])
        
        # Columns C-J: Event type dropdowns (Yes/No/Partial/N/A)
        for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Columns K-M: Format dropdowns
        for col_letter in ['K', 'L', 'M']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Columns N-O: Numeric inputs
        for col_letter in ['N', 'O']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column P: Storage tier dropdown
        apply_style(ws[f'P{data_row}'], styles['input_cell'])
        
        # Column Q: Protection mechanisms (text)
        apply_style(ws[f'Q{data_row}'], styles['input_cell'])
        
        # Column R: Event Types Completeness Formula
        ws[f'R{data_row}'] = f'=IF(A{data_row}="","",COUNTIF(C{data_row}:J{data_row},"*Yes*")/COUNTA(C{data_row}:J{data_row}))'
        ws[f'R{data_row}'].number_format = '0.0%'
        apply_style(ws[f'R{data_row}'], styles['formula_cell'])
        
        # Column S: Notes
        apply_style(ws[f'S{data_row}'], styles['input_cell'])
    
    # Data validations
    yes_no_dv = DataValidation(type="list",
        formula1='"\u2705 Yes,\u274C No,\u26A0\uFE0F Partial,➖ N/A"',
        allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    yes_no_dv.add('C9:J100')
    
    log_format_dv = DataValidation(type="list",
        formula1='"Syslog,CEF,JSON,EVTX,Custom,Unknown"',
        allow_blank=True)
    ws.add_data_validation(log_format_dv)
    log_format_dv.add('K9:K100')
    
    timestamp_format_dv = DataValidation(type="list",
        formula1='"ISO 8601,RFC 3339,Unix Epoch,Custom,Unknown"',
        allow_blank=True)
    ws.add_data_validation(timestamp_format_dv)
    timestamp_format_dv.add('L9:L100')
    
    timezone_dv = DataValidation(type="list",
        formula1='"UTC,Local,Unknown"',
        allow_blank=True)
    ws.add_data_validation(timezone_dv)
    timezone_dv.add('M9:M100')
    
    storage_tier_dv = DataValidation(type="list",
        formula1='"Hot,Warm,Cold,Unknown"',
        allow_blank=True)
    ws.add_data_validation(storage_tier_dv)
    storage_tier_dv.add('P9:P100')
    
    ws.freeze_panes = 'C8'

# ============================================================================
# SECTION 5: ASSESSMENT SHEET TEMPLATE FUNCTION
# ============================================================================

def create_assessment_sheet(ws, styles, sheet_config):
    """
    Generic function to create assessment sheets (Authentication, Authorization, etc.)
    
    sheet_config = {
        'title': 'Sheet title',
        'description': 'Instructions text',
        'assessment_columns': [list of tuples (col_letter, header, width)]
    }
    
    "Don't repeat yourself" - DRY principle. One function, many sheets!
    """
    
    # Header
    end_col = sheet_config['assessment_columns'][-1][0]
    ws.merge_cells(f'A1:{end_col}1')
    ws['A1'] = sheet_config['title']
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    # Instructions
    ws.merge_cells(f'A2:{end_col}2')
    ws['A2'] = sheet_config['description']
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Policy reference (Row 4)
    ws.merge_cells(f'A4:{end_col}4')
    ws['A4'] = f"Policy Requirement: {sheet_config.get('policy_ref', 'ISMS-POL-A.8.15-S2.1')}"
    ws['A4'].font = Font(bold=True, size=10, italic=True)
    ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Column headers (Row 7)
    row = 7
    for col_letter, header, width in sheet_config['assessment_columns']:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row (Row 8)
    row = 8
    if 'example_data' in sheet_config:
        for col_idx, value in enumerate(sheet_config['example_data'], start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-100)
    for data_row in range(9, 101):
        # Column A: System ID dropdown
        apply_style(ws[f'A{data_row}'], styles['input_cell'])
        
        # Column B: VLOOKUP System Name
        ws[f'B{data_row}'] = f'=IF(A{data_row}="","",IFERROR(VLOOKUP(A{data_row},\'System Inventory\'!A:B,2,FALSE),"Not Found"))'
        apply_style(ws[f'B{data_row}'], styles['formula_cell'])
        
        # Columns C through second-to-last: Input cells
        for col_letter, header, width in sheet_config['assessment_columns'][2:-3]:  # Skip A, B and last 3 cols
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column P (or configured): Compliance Score Formula
        score_col = sheet_config.get('score_column', 'P')
        # Get the range of assessment columns (C through O typically)
        first_assess_col = sheet_config['assessment_columns'][2][0]
        last_assess_col = chr(ord(score_col) - 3)  # 3 columns before score column
        
        ws[f'{score_col}{data_row}'] = f'=IF(A{data_row}="","",(COUNTIF({first_assess_col}{data_row}:{last_assess_col}{data_row},"*Yes*")/COUNTIF({first_assess_col}{data_row}:{last_assess_col}{data_row},"<>N/A"))*100)'
        ws[f'{score_col}{data_row}'].number_format = '0.0"%"'
        apply_style(ws[f'{score_col}{data_row}'], styles['formula_cell'])
        
        # Last two columns: Gap Description and Remediation Plan (text input)
        gap_col = chr(ord(score_col) + 1)
        remediation_col = chr(ord(score_col) + 2)
        apply_style(ws[f'{gap_col}{data_row}'], styles['input_cell'])
        apply_style(ws[f'{remediation_col}{data_row}'], styles['input_cell'])
    
    # Add data validations for Yes/No/Partial/N/A columns
    yes_no_cols = [col[0] for col in sheet_config['assessment_columns'][2:-3]]
    for col_letter in yes_no_cols:
        yes_no_dv = DataValidation(type="list",
            formula1='"\u2705 Yes,\u274C No,\u26A0\uFE0F Partial,➖ N/A"',
            allow_blank=True)
        ws.add_data_validation(yes_no_dv)
        yes_no_dv.add(f'{col_letter}9:{col_letter}100')
    
    ws.freeze_panes = 'C8'

# ============================================================================
# SECTION 6: AUTHENTICATION LOGGING ASSESSMENT SHEET
# ============================================================================

def create_authentication_logging_sheet(ws, styles):
    """
    Create Authentication Logging Assessment sheet.
    Per ISMS-POL-A.8.15-S2.1.2 - Authentication events.
    """
    
    config = {
        'title': 'AUTHENTICATION LOGGING ASSESSMENT',
        'description': 'Assess authentication event logging completeness per ISMS-POL-A.8.15-S2.1.2',
        'policy_ref': 'ISMS-POL-A.8.15-S2.1.2 (Authentication Events)',
        'score_column': 'P',
        'assessment_columns': [
            ("A", "System ID", 15),
            ("B", "System Name", 30),
            ("C", "Logs Successful Logins", 18),
            ("D", "Logs Failed Logins", 18),
            ("E", "Logs Account Lockouts", 18),
            ("F", "Logs Password Changes", 18),
            ("G", "Logs Session Start/End", 18),
            ("H", "Includes User ID", 15),
            ("I", "Includes Timestamp", 15),
            ("J", "Includes Source IP", 15),
            ("K", "Includes Auth Method", 15),
            ("L", "MFA Events Logged", 15),
            ("M", "SSO Events Logged", 15),
            ("N", "Service Account Auth", 20),
            ("O", "Privileged Auth Logged", 18),
            ("P", "Compliance Score", 18),
            ("Q", "Gap Description", 40),
            ("R", "Remediation Plan", 40),
        ],
        'example_data': [
            "SYS-001", "web-prod-01", "Yes", "Yes", "Yes", "Yes", "Yes",
            "Yes", "Yes", "Yes", "Yes", "Yes", "Yes", "N/A", "Yes",
            "100.0%", "", ""
        ]
    }
    
    create_assessment_sheet(ws, styles, config)

# ============================================================================
# SECTION 7: AUTHORIZATION & ACCESS LOGGING SHEET
# ============================================================================

def create_authorization_logging_sheet(ws, styles):
    """
    Create Authorization & Access Logging Assessment sheet.
    Per ISMS-POL-A.8.15-S2.1.3 - Authorization and access control events.
    """
    
    config = {
        'title': 'AUTHORIZATION & ACCESS LOGGING ASSESSMENT',
        'description': 'Assess access control event logging per ISMS-POL-A.8.15-S2.1.3',
        'policy_ref': 'ISMS-POL-A.8.15-S2.1.3 (Authorization Events)',
        'score_column': 'P',
        'assessment_columns': [
            ("A", "System ID", 15),
            ("B", "System Name", 30),
            ("C", "Logs Access Grants", 18),
            ("D", "Logs Access Denials", 18),
            ("E", "Logs Permission Changes", 18),
            ("F", "Logs Privilege Escalation", 18),
            ("G", "Logs Data Access (Sensitive)", 18),
            ("H", "Includes User ID", 15),
            ("I", "Includes Resource Accessed", 15),
            ("J", "Includes Action Type", 15),
            ("K", "Includes Outcome", 15),
            ("L", "Includes Reason (denied)", 15),
            ("M", "Includes Before/After State", 15),
            ("N", "Logs Group Membership", 20),
            ("O", "Logs Role Assignment", 18),
            ("P", "Compliance Score", 18),
            ("Q", "Gap Description", 40),
            ("R", "Remediation Plan", 40),
        ],
        'example_data': [
            "SYS-001", "web-prod-01", "Yes", "Yes", "Yes", "Yes", "Yes",
            "Yes", "Yes", "Yes", "Yes", "Partial", "Partial", "Yes", "Yes",
            "92.3%", "Before/After state partially logged", "Enhance logging"
        ]
    }
    
    create_assessment_sheet(ws, styles, config)

# ============================================================================
# SECTION 8: ADMINISTRATIVE ACTIVITY LOGGING SHEET
# ============================================================================

def create_administrative_activity_sheet(ws, styles):
    """
    Create Administrative Activity Logging Assessment sheet.
    Per ISMS-POL-A.8.15-S2.1.4 - Administrative actions.
    """
    
    config = {
        'title': 'ADMINISTRATIVE ACTIVITY LOGGING ASSESSMENT',
        'description': 'Assess administrative action logging per ISMS-POL-A.8.15-S2.1.4',
        'policy_ref': 'ISMS-POL-A.8.15-S2.1.4 (Administrative Actions)',
        'score_column': 'P',
        'assessment_columns': [
            ("A", "System ID", 15),
            ("B", "System Name", 30),
            ("C", "User Account Management", 18),
            ("D", "Group/Role Management", 18),
            ("E", "Configuration Changes", 18),
            ("F", "Security Policy Changes", 18),
            ("G", "Software Install/Uninstall", 18),
            ("H", "Service Start/Stop", 15),
            ("I", "Patch Application", 15),
            ("J", "Bulk Data Operations", 15),
            ("K", "Privileged Session Logging", 15),
            ("L", "Includes Administrator ID", 15),
            ("M", "Includes Before/After Values", 15),
            ("N", "Includes Timestamp", 20),
            ("O", "Includes Change Reason", 18),
            ("P", "Compliance Score", 18),
            ("Q", "Gap Description", 40),
            ("R", "Remediation Plan", 40),
        ],
        'example_data': [
            "SYS-001", "web-prod-01", "Yes", "Yes", "Yes", "Yes", "Partial",
            "Yes", "Partial", "No", "Yes", "Yes", "Partial", "Yes", "No",
            "76.9%", "Change reason not logged, bulk ops not logged", "Implement change logging"
        ]
    }
    
    create_assessment_sheet(ws, styles, config)

# ============================================================================
# SECTION 9: SECURITY EVENT LOGGING SHEET
# ============================================================================

def create_security_event_logging_sheet(ws, styles):
    """
    Create Security Event Logging Assessment sheet.
    Per ISMS-POL-A.8.15-S2.1.5 - Security tool and event logging.
    """
    
    config = {
        'title': 'SECURITY EVENT LOGGING ASSESSMENT',
        'description': 'Assess security tool and event logging per ISMS-POL-A.8.15-S2.1.5',
        'policy_ref': 'ISMS-POL-A.8.15-S2.1.5 (Security Events)',
        'score_column': 'P',
        'assessment_columns': [
            ("A", "System ID", 15),
            ("B", "System Name", 30),
            ("C", "Firewall Events", 18),
            ("D", "IDS/IPS Alerts", 18),
            ("E", "Anti-malware Events", 18),
            ("F", "DLP Events", 18),
            ("G", "Web Filtering Events", 18),
            ("H", "Email Gateway Events", 15),
            ("I", "EDR Events", 15),
            ("J", "Vulnerability Scan Results", 15),
            ("K", "Security Incident Events", 15),
            ("L", "Includes Severity Level", 15),
            ("M", "Includes Threat Indicators", 15),
            ("N", "Includes Response Actions", 20),
            ("O", "Automated Response Logged", 18),
            ("P", "Compliance Score", 18),
            ("Q", "Gap Description", 40),
            ("R", "Remediation Plan", 40),
        ],
        'example_data': [
            "FW-001", "perimeter-fw-01", "Yes", "Yes", "N/A", "N/A", "N/A",
            "N/A", "N/A", "N/A", "Yes", "Yes", "Yes", "Partial", "No",
            "100.0%", "Automated response not logged", "Enable response logging"
        ]
    }
    
    create_assessment_sheet(ws, styles, config)


# ============================================================================
# SECTION 10: APPLICATION & DATABASE LOGGING SHEET
# ============================================================================

def create_application_database_logging_sheet(ws, styles):
    """
    Create Application & Database Logging Assessment sheet.
    Per ISMS-POL-A.8.15-S2.1.7 - Application and database events.
    """
    
    config = {
        'title': 'APPLICATION & DATABASE LOGGING ASSESSMENT',
        'description': 'Assess application and database logging per ISMS-POL-A.8.15-S2.1.7',
        'policy_ref': 'ISMS-POL-A.8.15-S2.1.7 (Application & Database Events)',
        'score_column': 'P',
        'assessment_columns': [
            ("A", "System ID", 15),
            ("B", "System Name", 30),
            ("C", "Web App Access Logging", 18),
            ("D", "API Call Logging", 18),
            ("E", "Transaction Logging", 18),
            ("F", "Application Errors", 18),
            ("G", "Database Connections", 18),
            ("H", "Database Queries (Sensitive)", 15),
            ("I", "Schema Changes (DDL)", 15),
            ("J", "Permission Grants", 15),
            ("K", "Backup/Restore Operations", 15),
            ("L", "Includes User/App Identity", 15),
            ("M", "Includes Data Modified", 15),
            ("N", "Includes Query Text", 20),
            ("O", "Includes Row Count", 18),
            ("P", "Compliance Score", 18),
            ("Q", "Gap Description", 40),
            ("R", "Remediation Plan", 40),
        ],
        'example_data': [
            "DB-001", "prod-db-01", "N/A", "Yes", "Yes", "Yes", "Yes",
            "Yes", "Yes", "Yes", "Yes", "Yes", "Partial", "Partial", "Yes",
            "91.7%", "Query text truncated, data modified partially logged", "Enable full query logging"
        ]
    }
    
    create_assessment_sheet(ws, styles, config)


# ============================================================================
# SECTION 11: NETWORK DEVICE LOGGING SHEET
# ============================================================================

def create_network_device_logging_sheet(ws, styles):
    """
    Create Network Device Logging Assessment sheet.
    Per ISMS-POL-A.8.15-S2.1.8 - Network infrastructure logging.
    """
    
    config = {
        'title': 'NETWORK DEVICE LOGGING ASSESSMENT',
        'description': 'Assess network infrastructure logging per ISMS-POL-A.8.15-S2.1.8',
        'policy_ref': 'ISMS-POL-A.8.15-S2.1.8 (Network Events)',
        'score_column': 'P',
        'assessment_columns': [
            ("A", "Device ID", 15),
            ("B", "Device Name", 30),
            ("C", "Connection Logging", 18),
            ("D", "Rule Match Logging", 18),
            ("E", "Configuration Changes", 18),
            ("F", "Interface Up/Down Events", 18),
            ("G", "Routing Changes", 18),
            ("H", "VPN Session Logging", 15),
            ("I", "DHCP/DNS Events", 15),
            ("J", "Wireless Events", 15),
            ("K", "NAT Translations", 15),
            ("L", "Includes Source/Dest IP", 15),
            ("M", "Includes Ports/Protocols", 15),
            ("N", "Includes Action (allow/deny)", 20),
            ("O", "Includes Bytes Transferred", 18),
            ("P", "Compliance Score", 18),
            ("Q", "Gap Description", 40),
            ("R", "Remediation Plan", 40),
        ],
        'example_data': [
            "NET-001", "core-switch-01", "Yes", "Yes", "Yes", "Yes", "Yes",
            "N/A", "Yes", "N/A", "N/A", "Yes", "Yes", "Yes", "No",
            "100.0%", "Bytes transferred not logged", "Enable traffic volume logging"
        ]
    }
    
    create_assessment_sheet(ws, styles, config)

# ============================================================================
# SECTION 12: GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis_sheet(ws, styles):
    """
    Create Gap Analysis sheet - consolidated gaps from all assessments.
    
    "The gap between where you are and where you want to be is called a plan."
    - Unknown (probably an auditor)
    """
    
    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "GAP ANALYSIS - CONSOLIDATED FINDINGS"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    # Instructions
    ws.merge_cells('A2:L2')
    ws['A2'] = "Document all identified gaps from assessment sheets. Prioritize by risk and establish remediation plans with clear ownership and target dates."
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Column headers (Row 7)
    headers = [
        ("A", "Gap ID", 12),
        ("B", "System ID", 15),
        ("C", "System Name", 30),
        ("D", "Gap Category", 25),
        ("E", "Gap Description", 50),
        ("F", "Policy Requirement", 30),
        ("G", "Impact / Risk", 20),
        ("H", "Remediation Action", 50),
        ("I", "Responsible Party", 25),
        ("J", "Target Date", 15),
        ("K", "Status", 15),
        ("L", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row (Row 8)
    row = 8
    example_data = [
        "GAP-001", "SYS-042", "mail-server-01", "Event Type Not Logged",
        "Failed authentication attempts not logged",
        "ISMS-POL-A.8.15-S2.1.2", "High",
        "Configure mail server to log authentication failures",
        "System Admin Team", "31.03.2026", "Open",
        "Vendor support required for config change"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-100 = 92 gap rows)
    for data_row in range(9, 101):
        # Column A: Auto-generate Gap ID
        ws[f'A{data_row}'] = f'=IF(B{data_row}<>"","GAP-"&TEXT(ROW()-8,"000"),"")'
        apply_style(ws[f'A{data_row}'], styles['formula_cell'])
        
        # Column B: System ID dropdown (links to System Inventory)
        apply_style(ws[f'B{data_row}'], styles['input_cell'])
        
        # Column C: VLOOKUP System Name
        ws[f'C{data_row}'] = f'=IF(B{data_row}="","",IFERROR(VLOOKUP(B{data_row},\'System Inventory\'!A:B,2,FALSE),"Not Found"))'
        apply_style(ws[f'C{data_row}'], styles['formula_cell'])
        
        # Columns D-I: Input cells
        for col_letter in ['D', 'E', 'F', 'G', 'H', 'I']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column J: Target Date (with date format)
        apply_style(ws[f'J{data_row}'], styles['input_cell'])
        ws[f'J{data_row}'].number_format = 'DD.MM.YYYY'
        
        # Columns K-L: Status and Notes
        for col_letter in ['K', 'L']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
    
    # Data validations
    gap_category_dv = DataValidation(type="list",
        formula1='"Log Source Missing,Event Type Not Logged,Incomplete Fields,Format Non-Standard,Protection Inadequate,Other"',
        allow_blank=True)
    ws.add_data_validation(gap_category_dv)
    gap_category_dv.add('D9:D100')
    
    risk_dv = DataValidation(type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚫ Low"',
        allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add('G9:G100')
    
    status_dv = DataValidation(type="list",
        formula1='"\u274C Open,⏳ In Progress,\u2705 Resolved,⭕ Deferred"',
        allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add('K9:K100')
    
    # Summary section below data (Row 105+)
    row = 105
    ws.merge_cells(f'A{row}:L{row}')
    ws[f'A{row}'] = "GAP SUMMARY STATISTICS"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    summary_labels = [
        ("Total Gaps Identified:", f'=COUNTA(A9:A100)'),
        ("Open Gaps:", f'=COUNTIF(K9:K100,"*Open*")'),
        ("In Progress:", f'=COUNTIF(K9:K100,"*In Progress*")'),
        ("Resolved:", f'=COUNTIF(K9:K100,"*Resolved*")'),
        ("Critical Priority:", f'=COUNTIF(G9:G100,"*Critical*")'),
        ("High Priority:", f'=COUNTIF(G9:G100,"*High*")'),
        ("Overdue (Past Target Date):", f'=SUMPRODUCT((K9:K100<>"Resolved")*(J9:J100<TODAY())*(J9:J100<>""))'),
    ]
    
    for label, formula in summary_labels:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = formula
        apply_style(ws[f'B{row}'], styles['info_cell'])
        row += 1
    
    ws.freeze_panes = 'A8'

# ============================================================================
# SECTION 13: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register_sheet(ws, styles):
    """
    Create Evidence Register sheet - track evidence artifacts for audit.
    
    "In God we trust. All others must bring data." - W. Edwards Deming
    """
    
    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "EVIDENCE REGISTER"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    # Instructions
    ws.merge_cells('A2:J2')
    ws['A2'] = "Register all evidence artifacts collected during assessment. Include log samples, screenshots, configurations, and documentation."
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Column headers (Row 7)
    headers = [
        ("A", "Evidence ID", 15),
        ("B", "Evidence Type", 25),
        ("C", "Description", 40),
        ("D", "Related System(s)", 30),
        ("E", "Related Policy Req", 25),
        ("F", "File Name / Location", 40),
        ("G", "Collected By", 25),
        ("H", "Collection Date", 15),
        ("I", "Retention Period", 20),
        ("J", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row (Row 8)
    row = 8
    example_data = [
        "EVD-001", "Log Sample", "Authentication log sample from web server",
        "SYS-001 (web-prod-01)", "S2.1.2", "auth_logs_sample_20260106.txt",
        "SOC Analyst", "06.01.2026", "3 years", "Sample covers 24h period"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-108 = 100 evidence rows)
    for data_row in range(9, 109):
        # Column A: Auto-generate Evidence ID
        ws[f'A{data_row}'] = f'=IF(B{data_row}<>"","EVD-"&TEXT(ROW()-8,"000"),"")'
        apply_style(ws[f'A{data_row}'], styles['formula_cell'])
        
        # Columns B-G: Input cells
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column H: Collection Date (with date format)
        apply_style(ws[f'H{data_row}'], styles['input_cell'])
        ws[f'H{data_row}'].number_format = 'DD.MM.YYYY'
        
        # Columns I-J: Retention and Notes
        for col_letter in ['I', 'J']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
    
    # Data validations
    evidence_type_dv = DataValidation(type="list",
        formula1='"Log Sample,Configuration Screenshot,SIEM Query Result,Documentation,Policy Document,Other"',
        allow_blank=True)
    ws.add_data_validation(evidence_type_dv)
    evidence_type_dv.add('B9:B108')
    
    ws.freeze_panes = 'A8'

# ============================================================================
# SECTION 14: SUMMARY DASHBOARD SHEET (PART 1)
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """
    Create Summary Dashboard - executive view of assessment results.
    
    "What gets measured gets managed." - Peter Drucker
    Let's give management something worth managing!
    """
    
    # Main header
    ws.merge_cells('A1:H1')
    ws['A1'] = "LOGGING COMPLIANCE DASHBOARD"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:H2')
    ws['A2'] = "Executive Summary - ISO/IEC 27001:2022 Control A.8.15"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 25
    
    # ==================== SECTION 1: OVERALL COMPLIANCE SUMMARY ====================
    row = 4
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "1. OVERALL COMPLIANCE SUMMARY"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    # Create summary metrics table
    metrics = [
        ("Total Systems Assessed", "=COUNTA('System Inventory'!B9:B100)", "N/A"),
        ("Systems with Logging Enabled", "=COUNTIF('System Inventory'!O:O,\"Yes\")", "100%"),
        ("Systems Forwarding to SIEM", "=COUNTIF('System Inventory'!P:P,\"Yes\")", "100%"),
        ("Overall Compliance Rate", "=AVERAGE('Authentication Logging'!P9:P100,'Authorization & Access'!P9:P100,'Administrative Activity'!P9:P100,'Security Event Logging'!P9:P100,'Application & Database'!P9:P100,'Network Device Logging'!P9:P100)", ">95%"),
        ("Critical Systems Compliant (P1)", "=COUNTIFS('System Inventory'!I:I,\"P1-Critical\",'System Inventory'!Q:Q,\"Compliant\")/COUNTIF('System Inventory'!I:I,\"P1-Critical\")", "100%"),
        ("High Systems Compliant (P2)", "=COUNTIFS('System Inventory'!I:I,\"P2-High\",'System Inventory'!Q:Q,\"Compliant\")/COUNTIF('System Inventory'!I:I,\"P2-High\")", ">95%"),
        ("Medium Systems Compliant (P3)", "=COUNTIFS('System Inventory'!I:I,\"P3-Medium\",'System Inventory'!Q:Q,\"Compliant\")/COUNTIF('System Inventory'!I:I,\"P3-Medium\")", ">90%"),
    ]
    
    # Headers
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    row += 1
    for metric_name, formula, target in metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(bold=True, size=10)
        
        ws[f'B{row}'] = formula
        apply_style(ws[f'B{row}'], styles['formula_cell'])
        if "Rate" in metric_name or "Compliant" in metric_name:
            ws[f'B{row}'].number_format = '0.0%'
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Status column - conditional formula
        if "Rate" in metric_name or "Compliant" in metric_name:
            ws[f'D{row}'] = f'=IF(B{row}>={target[1:]},"✓ Met","✗ Not Met")'
        else:
            ws[f'D{row}'] = f'=IF(B{row}>0,"✓","✗")'
        apply_style(ws[f'D{row}'], styles['formula_cell'])
        
        row += 1
    
    # ==================== SECTION 2: EVENT TYPE COVERAGE ====================
    row += 2
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "2. EVENT TYPE COVERAGE"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    # Event type coverage table
    event_types = [
        ("Authentication", "=COUNTIF('Log Event Types by System'!C:C,\"Yes\")", ">95%"),
        ("Authorization", "=COUNTIF('Log Event Types by System'!D:D,\"Yes\")", ">95%"),
        ("Administrative", "=COUNTIF('Log Event Types by System'!E:E,\"Yes\")", ">95%"),
        ("Security Events", "=COUNTIF('Log Event Types by System'!F:F,\"Yes\")", "100%"),
        ("Application", "=COUNTIF('Log Event Types by System'!G:G,\"Yes\")", ">90%"),
        ("System", "=COUNTIF('Log Event Types by System'!H:H,\"Yes\")", ">85%"),
        ("Network", "=COUNTIF('Log Event Types by System'!I:I,\"Yes\")", ">85%"),
        ("Database", "=COUNTIF('Log Event Types by System'!J:J,\"Yes\")", ">90%"),
    ]
    
    # Headers
    ws[f'A{row}'] = "Event Type"
    ws[f'B{row}'] = "Systems Logging"
    ws[f'C{row}'] = "% Coverage"
    ws[f'D{row}'] = "Target"
    ws[f'E{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    row += 1
    for event_type, formula, target in event_types:
        ws[f'A{row}'] = event_type
        ws[f'A{row}'].font = Font(bold=True, size=10)
        
        ws[f'B{row}'] = formula
        apply_style(ws[f'B{row}'], styles['formula_cell'])
        
        # Calculate percentage
        ws[f'C{row}'] = f'=B{row}/COUNTA(\'System Inventory\'!B9:B100)'
        ws[f'C{row}'].number_format = '0.0%'
        apply_style(ws[f'C{row}'], styles['formula_cell'])
        
        ws[f'D{row}'] = target
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Status
        ws[f'E{row}'] = f'=IF(C{row}>={target[1:]},"✓ Met","✗ Not Met")'
        apply_style(ws[f'E{row}'], styles['formula_cell'])
        
        row += 1
    
    # Set column widths
    set_column_widths(ws, {
        'A': 35,
        'B': 20,
        'C': 15,
        'D': 15,
        'E': 15,
        'F': 20,
        'G': 20,
        'H': 20
    })

# ============================================================================
# SECTION 15: SUMMARY DASHBOARD SHEET (PART 2 - GAP SUMMARY)
# ============================================================================

def add_gap_summary_to_dashboard(ws, styles):
    """
    Add Gap Summary section to Dashboard sheet.
    Continues from the Event Type Coverage section.
    """
    
    # Start after Event Type Coverage section (approx row 30)
    row = 30
    
    # ==================== SECTION 3: GAP SUMMARY BY PRIORITY ====================
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "3. GAP SUMMARY BY PRIORITY"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    # Headers
    gap_headers = ["Priority", "Open Gaps", "In Progress", "Resolved", "Overdue", "% Complete"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row += 1
    priorities = ["Critical", "High", "Medium", "Low"]
    
    for priority in priorities:
        ws[f'A{row}'] = priority
        ws[f'A{row}'].font = Font(bold=True, size=10)
        
        # Open gaps
        ws[f'B{row}'] = f'=COUNTIFS(\'Gap Analysis\'!G:G,"*{priority}*",\'Gap Analysis\'!K:K,"*Open*")'
        apply_style(ws[f'B{row}'], styles['formula_cell'])
        
        # In Progress
        ws[f'C{row}'] = f'=COUNTIFS(\'Gap Analysis\'!G:G,"*{priority}*",\'Gap Analysis\'!K:K,"*In Progress*")'
        apply_style(ws[f'C{row}'], styles['formula_cell'])
        
        # Resolved
        ws[f'D{row}'] = f'=COUNTIFS(\'Gap Analysis\'!G:G,"*{priority}*",\'Gap Analysis\'!K:K,"*Resolved*")'
        apply_style(ws[f'D{row}'], styles['formula_cell'])
        
        # Overdue (Open or In Progress, past target date)
        ws[f'E{row}'] = f'=SUMPRODUCT((ISNUMBER(SEARCH("*{priority}*",\'Gap Analysis\'!G:G)))*((ISNUMBER(SEARCH("*Open*",\'Gap Analysis\'!K:K)))+(ISNUMBER(SEARCH("*In Progress*",\'Gap Analysis\'!K:K))))*(\'Gap Analysis\'!J:J<TODAY())*(\'Gap Analysis\'!J:J<>""))'
        apply_style(ws[f'E{row}'], styles['formula_cell'])
        
        # % Complete
        ws[f'F{row}'] = f'=IF((B{row}+C{row}+D{row})=0,0,D{row}/(B{row}+C{row}+D{row}))'
        ws[f'F{row}'].number_format = '0.0%'
        apply_style(ws[f'F{row}'], styles['formula_cell'])
        
        row += 1
    
    # Total row
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'] = f'=SUM({col}{row-4}:{col}{row-1})'
        apply_style(ws[f'{col}{row}'], styles['formula_cell'])
        ws[f'{col}{row}'].font = Font(bold=True)
    
    ws[f'F{row}'] = f'=IF((B{row}+C{row}+D{row})=0,0,D{row}/(B{row}+C{row}+D{row}))'
    ws[f'F{row}'].number_format = '0.0%'
    apply_style(ws[f'F{row}'], styles['formula_cell'])
    ws[f'F{row}'].font = Font(bold=True)
    
    # ==================== SECTION 4: TOP GAPS REQUIRING ATTENTION ====================
    row += 3
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "4. TOP GAPS REQUIRING ATTENTION"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "Top 10 gaps sorted by priority (Critical first) and target date (most overdue first)"
    ws[f'A{row}'].font = Font(italic=True, size=9)
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    row += 1
    top_gap_headers = ["Gap ID", "System", "Description", "Priority", "Target Date", "Status"]
    for col_idx, header in enumerate(top_gap_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row += 1
    note_row = row
    ws.merge_cells(f'A{note_row}:F{note_row}')
    ws[f'A{note_row}'] = "Note: Manually copy top 10 gaps from Gap Analysis sheet after sorting by Priority (desc) and Target Date (asc)"
    ws[f'A{note_row}'].font = Font(italic=True, size=9, color='666666')
    ws[f'A{note_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Reserve 10 rows for top gaps (manual entry or linked)
    for i in range(10):
        row += 1
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            apply_style(ws[f'{col}{row}'], styles['input_cell'])
    
    # ==================== SECTION 5: RECOMMENDATIONS ====================
    row += 2
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "5. RECOMMENDATIONS"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    recommendations = [
        "1. IMMEDIATE ACTIONS (Critical Gaps):",
        "   \u2022 Address all Critical priority gaps within 30 days",
        "   \u2022 Implement logging on all P1-Critical systems showing non-compliant status",
        "   \u2022 Enable SIEM forwarding for security event sources",
        "",
        "2. SHORT-TERM IMPROVEMENTS (30-90 days):",
        "   \u2022 Standardize log formats across similar system types",
        "   \u2022 Implement centralized authentication logging for all applications",
        "   \u2022 Complete missing administrative activity logging",
        "",
        "3. MEDIUM-TERM ENHANCEMENTS (3-6 months):",
        "   \u2022 Achieve >95% overall compliance rate across all event types",
        "   \u2022 Implement automated log analysis and alerting",
        "   \u2022 Establish quarterly log source review process",
        "",
        "4. RESOURCE REQUIREMENTS:",
        "   \u2022 Budget allocation for SIEM/log management capacity expansion",
        "   \u2022 Additional FTE or contractor support for gap remediation",
        "   \u2022 Training for system owners on logging best practices",
    ]
    
    for recommendation in recommendations:
        ws[f'A{row}'] = recommendation
        if recommendation.endswith(":"):
            ws[f'A{row}'].font = Font(bold=True, size=10)
        else:
            ws[f'A{row}'].font = Font(size=9)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
    
    # ==================== SECTION 6: ASSESSMENT METADATA ====================
    row += 2
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "6. ASSESSMENT METADATA"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    metadata = [
        ("Assessment Completed By:", "=\'Instructions & Legend\'!B6"),
        ("Assessment Date:", "=\'Instructions & Legend\'!B5"),
        ("Review Date:", "=TODAY()"),
        ("Next Assessment Due:", "=\'Instructions & Legend\'!B5+365"),
        ("Approval Status:", "See Approval & Sign-Off sheet"),
    ]
    
    for label, value in metadata:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True, size=10)
        
        ws[f'B{row}'] = value
        apply_style(ws[f'B{row}'], styles['info_cell'])
        if "Date" in label:
            ws[f'B{row}'].number_format = 'DD.MM.YYYY'
        
        row += 1


# ============================================================================
# SECTION 16: SUMMARY DASHBOARD CHARTS (OPTIONAL)
# ============================================================================

def add_charts_to_dashboard(ws):
    """
    Add charts to Summary Dashboard.
    
    Note: Chart generation is optional - can be added manually in Excel
    for better customization. This provides the basic structure.
    
    "A picture is worth a thousand words, but a good chart is worth
    a thousand meetings." - Unknown CFO
    """
    
    try:
        # Chart 1: Compliance Status Pie Chart
        pie = PieChart()
        pie.title = "Overall Compliance Status"
        pie.style = 10
        pie.height = 10
        pie.width = 15
        
        # Data for pie chart - would reference System Inventory compliance status
        # Note: This is simplified - actual implementation would need proper data ranges
        labels = Reference(ws, min_col=1, min_row=6, max_row=8)
        data = Reference(ws, min_col=2, min_row=5, max_row=8)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        
        ws.add_chart(pie, "J4")
        
        # Chart 2: Event Type Coverage Bar Chart
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Event Type Coverage %"
        bar.y_axis.title = 'Coverage %'
        bar.x_axis.title = 'Event Type'
        bar.height = 10
        bar.width = 15
        
        # Would reference Event Type Coverage section
        ws.add_chart(bar, "J20")
        
    except Exception as e:
        # Chart creation can fail - not critical, can be done manually
        logger.info(f"   Note: Chart auto-generation skipped (can be added manually): {e}")

# ============================================================================
# SECTION 17: APPROVAL & SIGN-OFF SHEET
# ============================================================================

def create_approval_signoff_sheet(ws, styles):
    """
    Create Approval & Sign-Off sheet with multi-level approval workflow.
    
    "Trust, but verify. And get signatures." - Modified Russian proverb
    """
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "APPROVAL & SIGN-OFF"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    # Instructions
    ws.merge_cells('A2:E2')
    ws['A2'] = "Multi-level approval workflow for Log Source Inventory Assessment"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 25
    
    # Approval table
    row = 4
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "APPROVAL WORKFLOW"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    # Headers
    approval_headers = ["Role", "Name", "Date", "Signature", "Status"]
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row += 1
    
    # Approval roles
    approval_roles = [
        ("System Owners (collective)", "[List names]", "", "_____", "☐ Reviewed"),
        ("Log Administrator", "[Name]", "", "_____", "☐ Reviewed"),
        ("SOC Lead", "[Name]", "", "_____", "☐ Reviewed"),
        ("Information Security Manager", "[Name]", "", "_____", "☐ Approved"),
        ("CISO", "[Name]", "", "_____", "☐ Approved"),
    ]
    
    for role, name, date, signature, status in approval_roles:
        ws[f'A{row}'] = role
        ws[f'A{row}'].font = Font(bold=True, size=10)
        
        ws[f'B{row}'] = name
        apply_style(ws[f'B{row}'], styles['input_cell'])
        
        ws[f'C{row}'] = date
        apply_style(ws[f'C{row}'], styles['input_cell'])
        ws[f'C{row}'].number_format = 'DD.MM.YYYY'
        
        ws[f'D{row}'] = signature
        apply_style(ws[f'D{row}'], styles['input_cell'])
        
        ws[f'E{row}'] = status
        ws[f'E{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
    
    # Acknowledgments checklist
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "ACKNOWLEDGMENTS CHECKLIST"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    acknowledgments = [
        "☐ Assessment completed per ISMS-POL-A.8.15-S2.1",
        "☐ All in-scope systems documented in System Inventory",
        "☐ Event type logging assessed for all systems",
        "☐ Gaps identified and prioritized",
        "☐ Remediation plan established with target dates",
        "☐ Resources committed for gap remediation",
        "☐ Evidence artifacts collected and registered",
        "☐ Next assessment date scheduled",
        "☐ Stakeholders informed of findings",
        "☐ Dashboard reviewed with management",
    ]
    
    for acknowledgment in acknowledgments:
        ws[f'A{row}'] = acknowledgment
        ws[f'A{row}'].font = Font(size=10)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
    
    # Comments section
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "APPROVAL COMMENTS / CONDITIONS"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 1
    ws.merge_cells(f'A{row}:E{row+8}')
    apply_style(ws[f'A{row}'], styles['input_cell'])
    ws.row_dimensions[row].height = 120
    
    # Set column widths
    set_column_widths(ws, {
        'A': 35,
        'B': 25,
        'C': 15,
        'D': 15,
        'E': 20
    })

# ============================================================================
# SECTION 18: CONDITIONAL FORMATTING
# ============================================================================

def apply_conditional_formatting(wb):
    """
    Apply conditional formatting rules across all relevant sheets.
    
    "Color is a power which directly influences the soul." - Kandinsky
    Let's use it to influence compliance behavior!
    """
    
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    
    # Define fills for conditional formatting
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    
    # ==================== SYSTEM INVENTORY SHEET ====================
    ws = wb['System Inventory']
    
    # Compliance Status column (Q)
    ws.conditional_formatting.add('Q9:Q100',
        CellIsRule(operator='equal', formula=['"Compliant"'], stopIfTrue=True, fill=green_fill))
    ws.conditional_formatting.add('Q9:Q100',
        CellIsRule(operator='equal', formula=['"Partial"'], stopIfTrue=True, fill=yellow_fill))
    ws.conditional_formatting.add('Q9:Q100',
        CellIsRule(operator='equal', formula=['"Non-Compliant"'], stopIfTrue=True, fill=red_fill))
    
    # ==================== LOG EVENT TYPES SHEET ====================
    ws = wb['Log Event Types by System']
    
    # Event Types Completeness (R) - percentage-based
    ws.conditional_formatting.add('R9:R100',
        CellIsRule(operator='greaterThanOrEqual', formula=['0.9'], stopIfTrue=True, fill=green_fill))
    ws.conditional_formatting.add('R9:R100',
        CellIsRule(operator='between', formula=['0.7', '0.89'], stopIfTrue=True, fill=yellow_fill))
    ws.conditional_formatting.add('R9:R100',
        CellIsRule(operator='lessThan', formula=['0.7'], stopIfTrue=True, fill=red_fill))
    
    # ==================== ASSESSMENT SHEETS ====================
    assessment_sheets = [
        'Authentication Logging',
        'Authorization & Access',
        'Administrative Activity',
        'Security Event Logging',
        'Application & Database',
        'Network Device Logging'
    ]
    
    for sheet_name in assessment_sheets:
        ws = wb[sheet_name]
        
        # Compliance Score column (P) - percentage values
        ws.conditional_formatting.add('P9:P100',
            CellIsRule(operator='greaterThanOrEqual', formula=['95'], stopIfTrue=True, fill=green_fill))
        ws.conditional_formatting.add('P9:P100',
            CellIsRule(operator='between', formula=['80', '94'], stopIfTrue=True, fill=yellow_fill))
        ws.conditional_formatting.add('P9:P100',
            CellIsRule(operator='lessThan', formula=['80'], stopIfTrue=True, fill=red_fill))
    
    # ==================== GAP ANALYSIS SHEET ====================
    ws = wb['Gap Analysis']
    
    # Priority column (G)
    ws.conditional_formatting.add('G9:G100',
        CellIsRule(operator='equal', formula=['"Critical"'], stopIfTrue=True, fill=red_fill))
    ws.conditional_formatting.add('G9:G100',
        CellIsRule(operator='equal', formula=['"High"'], stopIfTrue=True, 
                   fill=PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')))
    ws.conditional_formatting.add('G9:G100',
        CellIsRule(operator='equal', formula=['"Medium"'], stopIfTrue=True, fill=yellow_fill))
    ws.conditional_formatting.add('G9:G100',
        CellIsRule(operator='equal', formula=['"Low"'], stopIfTrue=True, fill=green_fill))
    
    # Status column (K)
    ws.conditional_formatting.add('K9:K100',
        CellIsRule(operator='equal', formula=['"Open"'], stopIfTrue=True, fill=red_fill))
    ws.conditional_formatting.add('K9:K100',
        CellIsRule(operator='equal', formula=['"In Progress"'], stopIfTrue=True, fill=yellow_fill))
    ws.conditional_formatting.add('K9:K100',
        CellIsRule(operator='equal', formula=['"Resolved"'], stopIfTrue=True, fill=green_fill))
    ws.conditional_formatting.add('K9:K100',
        CellIsRule(operator='equal', formula=['"Deferred"'], stopIfTrue=True, fill=gray_fill))
    
    # Target Date column (J) - highlight overdue dates
    # Note: Date-based conditional formatting needs special handling in openpyxl
    # This is a simplified version
    from openpyxl.formatting.rule import FormulaRule
    ws.conditional_formatting.add('J9:J100',
        FormulaRule(formula=['AND(J9<TODAY(),K9<>"Resolved")'], stopIfTrue=True, fill=red_fill))
    
# ============================================================================
# SECTION 19: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """
    Main execution function - orchestrates workbook creation.
    
    "The secret to getting ahead is getting started." - Mark Twain
    Let's get this ISMS implementation started properly!
    """
    
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.15.1 - Log Source Inventory Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.15: Logging")
    logger.info("=" * 78)
    logger.info("")
    
    # Create workbook and styles
    logger.info("[1/15] Creating workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    
    # Create each sheet
    logger.info("[2/15] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    
    logger.info("[3/15] Creating System Inventory...")
    create_system_inventory_sheet(wb["System Inventory"], styles)
    
    logger.info("[4/15] Creating Log Event Types by System...")
    create_log_event_types_sheet(wb["Log Event Types by System"], styles)
    
    logger.info("[5/15] Creating Authentication Logging Assessment...")
    create_authentication_logging_sheet(wb["Authentication Logging"], styles)
    
    logger.info("[6/15] Creating Authorization & Access Logging Assessment...")
    create_authorization_logging_sheet(wb["Authorization & Access"], styles)
    
    logger.info("[7/15] Creating Administrative Activity Logging Assessment...")
    create_administrative_activity_sheet(wb["Administrative Activity"], styles)
    
    logger.info("[8/15] Creating Security Event Logging Assessment...")
    create_security_event_logging_sheet(wb["Security Event Logging"], styles)
    
    logger.info("[9/15] Creating Application & Database Logging Assessment...")
    create_application_database_logging_sheet(wb["Application & Database"], styles)
    
    logger.info("[10/15] Creating Network Device Logging Assessment...")
    create_network_device_logging_sheet(wb["Network Device Logging"], styles)
    
    logger.info("[11/15] Creating Gap Analysis...")
    create_gap_analysis_sheet(wb["Gap Analysis"], styles)
    
    logger.info("[12/15] Creating Evidence Register...")
    create_evidence_register_sheet(wb["Evidence Register"], styles)
    
    logger.info("[13/15] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)
    add_gap_summary_to_dashboard(wb["Summary Dashboard"], styles)
    try:
        add_charts_to_dashboard(wb["Summary Dashboard"])
    except Exception as e:
        logger.info(f"    Note: Charts can be added manually in Excel")
    
    logger.info("[14/15] Creating Approval & Sign-Off...")
    create_approval_signoff_sheet(wb["Approval & Sign-Off"], styles)
    
    logger.info("[15/15] Applying conditional formatting...")
    apply_conditional_formatting(wb)
    
    # Generate filename with today's date
    filename = f"ISMS-IMP-A.8.15.1_Log_Source_Inventory_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    logger.info("")
    logger.info("Saving workbook...")
    wb.save(filename)
    
    logger.info("")
    logger.info("=" * 78)
    logger.info("\u2705 SUCCESS: Workbook generated successfully!")
    logger.info("=" * 78)
    logger.info("")
    logger.info(f"📄 Filename: {filename}")
    logger.info(f"📊 File size: ~800 KB - 1.5 MB (estimated)")
    logger.info("")
    logger.info("Workbook Structure:")
    logger.info("  ✓ Sheet 1:  Instructions & Legend")
    logger.info("  ✓ Sheet 2:  System Inventory (92 data rows)")
    logger.info("  ✓ Sheet 3:  Log Event Types by System (92 data rows)")
    logger.info("  ✓ Sheet 4:  Authentication Logging Assessment (92 rows)")
    logger.info("  ✓ Sheet 5:  Authorization & Access Logging (92 rows)")
    logger.info("  ✓ Sheet 6:  Administrative Activity Logging (92 rows)")
    logger.info("  ✓ Sheet 7:  Security Event Logging (92 rows)")
    logger.info("  ✓ Sheet 8:  Application & Database Logging (92 rows)")
    logger.info("  ✓ Sheet 9:  Network Device Logging (92 rows)")
    logger.info("  ✓ Sheet 10: Gap Analysis (92 gap rows)")
    logger.info("  ✓ Sheet 11: Evidence Register (100 evidence rows)")
    logger.info("  ✓ Sheet 12: Summary Dashboard (with KPIs and gap summary)")
    logger.info("  ✓ Sheet 13: Approval & Sign-Off (multi-level workflow)")
    logger.info("")
    logger.info("Features:")
    logger.info("  ✓ Auto-generated IDs (System, Gap, Evidence)")
    logger.info("  ✓ VLOOKUP formulas for system names")
    logger.info("  ✓ Compliance score calculations")
    logger.info("  ✓ Data validation dropdowns")
    logger.info("  ✓ Conditional formatting (Green/Yellow/Red status)")
    logger.info("  ✓ Date format: DD.MM.YYYY")
    logger.info("  ✓ Frozen panes for easy navigation")
    logger.info("")
    logger.info("Next Steps:")
    logger.info("  1. Open the workbook in Excel/LibreOffice")
    logger.info("  2. Fill in yellow-highlighted cells (user input required)")
    logger.info("  3. Complete System Inventory sheet first")
    logger.info("  4. Then complete assessment sheets for each system")
    logger.info("  5. Review Summary Dashboard for compliance status")
    logger.info("  6. Document gaps in Gap Analysis sheet")
    logger.info("  7. Collect evidence artifacts in Evidence Register")
    logger.info("  8. Obtain approvals in Approval & Sign-Off sheet")
    logger.info("")
    logger.info("Estimated Completion Time: 4-8 hours (for 50-100 systems)")
    logger.info("")
    logger.info("═" * 78)
    logger.info("'Science is the belief in the ignorance of experts.' - Richard Feynman")
    logger.info("Don't cargo cult this assessment - actually read your logs!")
    logger.info("═" * 78)
    logger.info("")


if __name__ == "__main__":
    main()


# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
