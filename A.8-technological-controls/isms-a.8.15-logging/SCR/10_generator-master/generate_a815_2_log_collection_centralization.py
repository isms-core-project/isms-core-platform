#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.15.2 - Log Collection & Centralization Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.15: Logging
Assessment Domain 2 of 4: Log Collection and Centralization Infrastructure

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific log collection infrastructure, SIEM architecture,
and centralization requirements.

Key customization areas:
1. Log collection methods and protocols (match your infrastructure)
2. SIEM/log management platform capabilities (specific to your tooling)
3. Network architecture and log forwarding paths (based on your topology)
4. Collection reliability thresholds (adapt to your SLA requirements)
5. Compliance scoring criteria (aligned with your operational targets)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.15 Logging Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
log collection mechanisms and centralization infrastructure across the
organization's logging ecosystem.

**Purpose:**
Enables systematic assessment of log forwarding, collection reliability, and
centralization against ISO 27001:2022 Control A.8.15 requirements, supporting
evidence-based validation of log aggregation capabilities.

**Assessment Scope:**
- SIEM/log management platform infrastructure
- Log collection agents and forwarders
- Syslog collection infrastructure
- API-based log collection mechanisms
- Cloud platform log forwarding
- Collection reliability and completeness
- Network paths and bandwidth adequacy
- Buffering and queuing mechanisms
- Collection failure detection and alerting
- Gap analysis for uncollected log sources
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and collection standards
2. SIEM Platform Assessment - Central logging infrastructure evaluation
3. Collection Methods - Agent, syslog, API assessment by system
4. Network & Transport - Network paths and bandwidth assessment
5. Reliability & Completeness - Collection success rate evaluation
6. Failure Detection - Alerting and monitoring assessment
7. Buffer & Queue Management - Handling of collection failures
8. Cloud Log Collection - Cloud platform forwarding assessment
9. Gap Analysis - Log sources not forwarded to SIEM
10. Evidence Register - Audit evidence tracking and documentation
11. Summary Dashboard - Collection metrics and KPIs
12. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with dropdown lists for consistency
- Conditional formatting for collection status visualization
- Automated gap identification for uncollected log sources
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with log source inventory (A.8.15.1)

**Integration:**
This assessment feeds into the A.8.15.5 Compliance Dashboard, which
consolidates data from all four logging assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a815_2_log_collection_centralization.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a815_2_log_collection_centralization.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a815_2_log_collection_centralization.py --date 20250124

Output:
    File: ISMS_A_8_15_2_Log_Collection_Centralization_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review log source inventory from A.8.15.1 assessment
    2. Document log collection infrastructure and SIEM platform
    3. Complete collection method assessments for each log source
    4. Validate collection reliability and completeness
    5. Review gap analysis for log sources not forwarded
    6. Define remediation actions for collection failures
    7. Collect and link audit evidence (SIEM configs, collection stats)
    8. Obtain stakeholder approvals
    9. Feed results into A.8.15.5 Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.15
Assessment Domain:    2 of 4 (Log Collection and Centralization Infrastructure)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.15: Logging Policy (Governance)
    - ISMS-IMP-A.8.15.1: Log Source Inventory Assessment (Domain 1)
    - ISMS-IMP-A.8.15.2: Log Collection & Centralization Implementation Guide
    - ISMS-IMP-A.8.15.3: Log Protection & Retention Assessment (Domain 3)
    - ISMS-IMP-A.8.15.4: Log Analysis & Review Assessment (Domain 4)
    - ISMS-IMP-A.8.15.5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 24.01.2025
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.15.2 specification
    - Supports comprehensive log collection and centralization evaluation
    - Integrated with A.8.15.5 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Collection Architecture:**
Log collection reliability is critical for security monitoring effectiveness.
Ensure collection mechanisms handle network failures, buffering requirements,
and high-volume log generation scenarios. Review SIEM vendor best practices
and sizing guidelines.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of log forwarding for all critical systems
and demonstration of collection reliability monitoring.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- SIEM platform architecture and sizing
- Network topology and collection paths
- Security control implementation details
- Collection failure patterns and gaps

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Monthly: Check collection success rates and failure patterns
- Quarterly: Review new log sources added to infrastructure
- Semi-annually: Validate SIEM capacity and performance
- Annually: Complete reassessment of collection infrastructure
- Ad-hoc: When infrastructure changes or collection issues detected

**Quality Assurance:**
Have SIEM administrators and security operations engineers validate assessments
before using results for compliance reporting or remediation decisions.

**Performance Monitoring:**
Collection reliability should be continuously monitored. Establish baselines
for collection success rates (target: >99%) and alert on deviations. Review
collection statistics regularly to identify systemic issues.

**Regulatory Alignment:**
Ensure log collection meets regulatory requirements:
- PCI DSS: Centralized logging and monitoring requirements
- GDPR: Audit trail completeness for data processing activities
- HIPAA: Audit trail collection and retention requirements
- SOX: Financial system audit log collection

**Scalability Planning:**
Document SIEM platform capacity limits and growth projections. Plan for
capacity expansion before reaching 70% of platform limits. Consider log
volume trends when assessing collection infrastructure adequacy.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •

# Document identification constants
DOCUMENT_ID = "ISMS-IMP-A.8.15.2"
CONTROL_REF = "ISO/IEC 27001:2022 - Control A.8.15: Logging"

from openpyxl.chart import BarChart, LineChart, Reference


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 14 sheets - comprehensive log collection assessment
    sheets = [
        "Instructions & Legend",
        "SIEM Platform Details",
        "Log Forwarder Inventory",
        "Collection Reliability",
        "Integration Architecture",
        "SIEM Storage & Capacity",
        "Log Parsing & Normalization",
        "SIEM Performance Metrics",
        "Data Quality Assessment",
        "Encryption & Authentication",
        "Gap Analysis & Remediation",
        "Evidence Register",
        "Summary Dashboard",
        "Approval & Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles.
    
    "Simplicity is prerequisite for reliability." - Dijkstra
    Same styles as A.8.15.1 for consistency across workbooks.
    """
    return {
        'header_main': {
            'font': {'name': 'Calibri', 'size': 14, 'bold': True, 'color': 'FFFFFF'},
            'fill': {'start_color': '003366', 'end_color': '003366', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'header_sub': {
            'font': {'name': 'Calibri', 'size': 11, 'bold': True, 'color': 'FFFFFF'},
            'fill': {'start_color': '4472C4', 'end_color': '4472C4', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'column_header': {
            'font': {'name': 'Calibri', 'size': 10, 'bold': True, 'color': '000000'},
            'fill': {'start_color': 'D9D9D9', 'end_color': 'D9D9D9', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'input_cell': {
            'fill': {'start_color': 'FFFFCC', 'end_color': 'FFFFCC', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'left', 'vertical': 'top', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'example_cell': {
            'fill': {'start_color': 'E7E6E6', 'end_color': 'E7E6E6', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'italic': True, 'color': '666666'},
            'alignment': {'horizontal': 'left', 'vertical': 'top'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'formula_cell': {
            'fill': {'start_color': 'E7E6E6', 'end_color': 'E7E6E6', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'center', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'info_cell': {
            'fill': {'start_color': 'E7E6E6', 'end_color': 'E7E6E6', 'fill_type': 'solid'},
            'font': {'name': 'Calibri', 'size': 10, 'color': '000000'},
            'alignment': {'horizontal': 'left', 'vertical': 'top', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'section_header': {
            'font': {'name': 'Calibri', 'size': 11, 'bold': True, 'color': 'FFFFFF'},
            'fill': {'start_color': '4472C4', 'end_color': '4472C4', 'fill_type': 'solid'},
            'alignment': {'horizontal': 'left', 'vertical': 'center'},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        }
    }


def apply_style(cell, style_template):
    """Apply a style template to a cell."""
    if 'font' in style_template:
        cell.font = Font(**style_template['font'])
    if 'fill' in style_template:
        cell.fill = PatternFill(**style_template['fill'])
    if 'alignment' in style_template:
        cell.alignment = Alignment(**style_template['alignment'])
    if 'border' in style_template:
        cell.border = Border(
            left=Side(style=style_template['border'].get('left', 'thin')),
            right=Side(style=style_template['border'].get('right', 'thin')),
            top=Side(style=style_template['border'].get('top', 'thin')),
            bottom=Side(style=style_template['border'].get('bottom', 'thin'))
        )


def set_column_widths(ws, widths):
    """Set column widths from a dictionary {column_letter: width}."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

# ============================================================================
# SECTION 2: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    
    # Main header
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{DOCUMENT_ID}  -  Log Collection & Centralization Assessment\n{CONTROL_REF}"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    # Document Information Block
    row = 4
    info_fields = [
        ("Document ID:", "ISMS-IMP-A.8.15.2"),
        ("Assessment Area:", "Log Collection Infrastructure and SIEM Integration"),
        ("Related Policy:", "ISMS-POL-A.8.15-S2.1, S2.2"),
        ("Version:", "1.0"),
        ("Assessment Date:", "[USER INPUT - DD.MM.YYYY]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Annual (full assessment), Quarterly (metrics update)"),
    ]
    
    for label, value in info_fields:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = value
        
        if "[USER INPUT" in value:
            apply_style(ws[f'B{row}'], styles['input_cell'])
            if "Date" in label:
                ws[f'B{row}'].number_format = 'DD.MM.YYYY'
        else:
            apply_style(ws[f'B{row}'], styles['info_cell'])
        
        row += 1
    
    # How to Use This Workbook
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "HOW TO USE THIS WORKBOOK"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 1
    instructions = [
        "1. Document SIEM/log management platform details",
        "2. Catalog all log forwarders and collectors",
        "3. Assess collection reliability and performance",
        "4. Document integration architecture",
        "5. Review capacity and scalability",
        "6. Identify collection gaps and failures",
        "7. Review Summary Dashboard for overall health",
        "8. Document gaps and remediation plans",
        "9. Obtain approvals in Approval & Sign-Off sheet"
    ]
    
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        row += 1
    
    # Legend (same as A.8.15.1)
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "LEGEND - CELL COLOR CODING"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 1
    legend_items = [
        ("Yellow (FFFFCC)", "User input required"),
        ("Green (C6EFCE)", "Compliant / Healthy"),
        ("Red (FFC7CE)", "Non-compliant / Critical"),
        ("Gray (E7E6E6)", "Auto-calculated / Reference data"),
        ("Blue (4472C4)", "Instructions / Headers"),
    ]
    
    for color_desc, meaning in legend_items:
        ws[f'A{row}'] = color_desc
        ws[f'B{row}'] = meaning
        
        if "Yellow" in color_desc:
            ws[f'A{row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        elif "Green" in color_desc:
            ws[f'A{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif "Red" in color_desc:
            ws[f'A{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif "Gray" in color_desc:
            ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        elif "Blue" in color_desc:
            ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            ws[f'A{row}'].font = Font(color='FFFFFF', bold=True)
        
        ws[f'A{row}'].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        row += 1
    
    # Key Definitions
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "KEY DEFINITIONS"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 1
    definitions = [
        ("SIEM", "Security Information and Event Management platform"),
        ("Log Forwarder", "Agent that collects and forwards logs (Splunk UF, Fluentd, etc.)"),
        ("Indexer", "SIEM component that processes and stores logs"),
        ("Search Head", "SIEM component for querying logs"),
        ("Collection Reliability", "% of expected logs successfully received"),
        ("Parse Success Rate", "% of logs successfully parsed into fields"),
        ("Indexing Lag", "Time delay between log generation and searchability"),
        ("Hot/Warm/Cold", "Tiered storage based on age and access frequency"),
        ("EPS", "Events Per Second - log ingestion rate"),
        ("Data Quality", "Completeness, accuracy, and consistency of log data"),
    ]
    
    for term, definition in definitions:
        ws[f'A{row}'] = term
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = definition
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        row += 1
    
    set_column_widths(ws, {'A': 25, 'B': 60, 'C': 15, 'D': 15, 'E': 15, 'F': 15})
    ws.freeze_panes = 'A3'

# ============================================================================
# SECTION 3: SIEM PLATFORM DETAILS SHEET
# ============================================================================

def create_siem_platform_sheet(ws, styles):
    """Create SIEM Platform Details sheet."""
    
    # Header
    ws.merge_cells('A1:N1')
    ws['A1'] = "SIEM PLATFORM DETAILS"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:N2')
    ws['A2'] = "Document SIEM/log management platform architecture and components"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        ("A", "Platform Component", 25),
        ("B", "Product/Vendor", 25),
        ("C", "Version", 15),
        ("D", "Hostname/FQDN", 30),
        ("E", "IP Address", 15),
        ("F", "Role", 20),
        ("G", "OS/Platform", 20),
        ("H", "CPU Cores", 10),
        ("I", "Memory (GB)", 15),
        ("J", "Storage Capacity (TB)", 20),
        ("K", "Storage Used (TB)", 20),
        ("L", "Storage % Used", 15),
        ("M", "Availability", 15),
        ("N", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "SIEM Core", "Splunk Enterprise", "9.1.3", "siem-core-01.example.com",
        "10.10.1.10", "Primary", "Red Hat Linux 8", "32", "256", "50", "32",
        "64%", "Active", "Primary indexer cluster node"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-30)
    for data_row in range(9, 31):
        # Columns A-K, M-N: Input cells
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'M', 'N']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column L: Storage % Used Formula
        ws[f'L{data_row}'] = f'=IF(J{data_row}=0,0,K{data_row}/J{data_row})'
        ws[f'L{data_row}'].number_format = '0.0%'
        apply_style(ws[f'L{data_row}'], styles['formula_cell'])
    
    # Data validations
    component_dv = DataValidation(type="list",
        formula1='"SIEM Core,Indexer,Search Head,Forwarder Management,Storage,API Gateway,Other"',
        allow_blank=True)
    ws.add_data_validation(component_dv)
    component_dv.add('A9:A30')
    
    vendor_dv = DataValidation(type="list",
        formula1='"Splunk,QRadar,Sentinel,LogRhythm,ELK Stack,Graylog,Custom,Other"',
        allow_blank=True)
    ws.add_data_validation(vendor_dv)
    vendor_dv.add('B9:B30')
    
    role_dv = DataValidation(type="list",
        formula1='"Primary,Secondary,DR,Dev/Test"',
        allow_blank=True)
    ws.add_data_validation(role_dv)
    role_dv.add('F9:F30')
    
    availability_dv = DataValidation(type="list",
        formula1='"Active,Standby,Offline,Degraded"',
        allow_blank=True)
    ws.add_data_validation(availability_dv)
    availability_dv.add('M9:M30')
    
    ws.freeze_panes = 'A8'

# ============================================================================
# SECTION 4: LOG FORWARDER INVENTORY SHEET
# ============================================================================

def create_log_forwarder_inventory_sheet(ws, styles):
    """Create Log Forwarder Inventory sheet."""
    
    # Header
    ws.merge_cells('A1:R1')
    ws['A1'] = "LOG FORWARDER INVENTORY"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:R2')
    ws['A2'] = "Catalog all log forwarders/collectors deployed across infrastructure"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        ("A", "Forwarder ID", 15),
        ("B", "Forwarder Type", 20),
        ("C", "Version", 12),
        ("D", "Installed On System", 30),
        ("E", "System Hostname", 30),
        ("F", "Destination SIEM", 25),
        ("G", "Transport Protocol", 18),
        ("H", "Port", 10),
        ("I", "Encryption", 15),
        ("J", "Buffer Enabled", 15),
        ("K", "Buffer Size (MB)", 18),
        ("L", "Install Date", 15),
        ("M", "Last Updated", 15),
        ("N", "Status", 15),
        ("O", "Events/Day", 20),
        ("P", "Data/Day (MB)", 20),
        ("Q", "Last Health Check", 15),
        ("R", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "FWD-001", "Splunk UF", "9.1.3", "SYS-001", "web-prod-01",
        "Primary", "Syslog/TLS", "9997", "Yes (TLS 1.3)", "Yes", "100",
        "15.06.2025", "01.12.2025", "Running", "250000", "150",
        "06.01.2026", "Forwarding web server logs"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-200)
    for data_row in range(9, 201):
        # Column A: Auto-generate Forwarder ID
        ws[f'A{data_row}'] = f'=IF(B{data_row}<>"","FWD-"&TEXT(ROW()-8,"000"),"")'
        apply_style(ws[f'A{data_row}'], styles['formula_cell'])
        
        # All other columns: Input cells
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
            
            # Date format for date columns
            if col_letter in ['L', 'M', 'Q']:
                ws[f'{col_letter}{data_row}'].number_format = 'DD.MM.YYYY'
    
    # Data validations
    forwarder_type_dv = DataValidation(type="list",
        formula1='"Splunk UF,Fluentd,Logstash,rsyslog,syslog-ng,Filebeat,Winlogbeat,Windows Event Forwarder,Cloud Connector,Other"',
        allow_blank=True)
    ws.add_data_validation(forwarder_type_dv)
    forwarder_type_dv.add('B9:B200')
    
    destination_dv = DataValidation(type="list",
        formula1='"Primary,Secondary,Both,Load Balanced"',
        allow_blank=True)
    ws.add_data_validation(destination_dv)
    destination_dv.add('F9:F200')
    
    protocol_dv = DataValidation(type="list",
        formula1='"Syslog/TLS,Syslog/TCP,Syslog/UDP,HTTPS,gRPC,Proprietary"',
        allow_blank=True)
    ws.add_data_validation(protocol_dv)
    protocol_dv.add('G9:G200')
    
    encryption_dv = DataValidation(type="list",
        formula1='"Yes (TLS 1.3),Yes (TLS 1.2),No,N/A"',
        allow_blank=True)
    ws.add_data_validation(encryption_dv)
    encryption_dv.add('I9:I200')
    
    buffer_dv = DataValidation(type="list",
        formula1='"Yes,No"',
        allow_blank=True)
    ws.add_data_validation(buffer_dv)
    buffer_dv.add('J9:J200')
    
    status_dv = DataValidation(type="list",
        formula1='"Running,Stopped,Error,Unknown"',
        allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add('N9:N200')
    
    ws.freeze_panes = 'C8'

# ============================================================================
# SECTION 5: COLLECTION RELIABILITY SHEET
# ============================================================================

def create_collection_reliability_sheet(ws, styles):
    """Create Collection Reliability sheet."""
    
    # Header
    ws.merge_cells('A1:P1')
    ws['A1'] = "COLLECTION RELIABILITY"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:P2')
    ws['A2'] = "Track log collection reliability metrics over assessment period (last 30 days)"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        ("A", "Source System ID", 15),
        ("B", "Source System Name", 30),
        ("C", "Expected Events/Day", 20),
        ("D", "Actual Events/Day", 20),
        ("E", "Collection Rate %", 18),
        ("F", "Status", 15),
        ("G", "Last Event Received", 18),
        ("H", "Gap Detected", 15),
        ("I", "Gap Start Time", 18),
        ("J", "Gap End Time", 18),
        ("K", "Gap Duration (hours)", 20),
        ("L", "Gap Reason", 30),
        ("M", "Resolution Actions", 40),
        ("N", "Resolved By", 20),
        ("O", "Resolved Date", 15),
        ("P", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "SYS-001", "web-prod-01", "250000", "248000", "99.2%", "Compliant",
        "06.01.2026 23:55", "No", "", "", "", "", "", "", "", "Healthy collection"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-200)
    for data_row in range(9, 201):
        # Column A: Input (System ID)
        apply_style(ws[f'A{data_row}'], styles['input_cell'])
        
        # Column B: VLOOKUP System Name (from IMP-1)
        ws[f'B{data_row}'] = f'=IF(A{data_row}="","",A{data_row}&" (from IMP-1)")'
        apply_style(ws[f'B{data_row}'], styles['formula_cell'])
        
        # Columns C-D: Input (numbers)
        for col_letter in ['C', 'D']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column E: Collection Rate Formula
        ws[f'E{data_row}'] = f'=IF(C{data_row}=0,0,D{data_row}/C{data_row})'
        ws[f'E{data_row}'].number_format = '0.0%'
        apply_style(ws[f'E{data_row}'], styles['formula_cell'])
        
        # Column F: Status Formula
        ws[f'F{data_row}'] = f'=IF(A{data_row}="","",IF(E{data_row}>=0.95,"Compliant",IF(E{data_row}>=0.8,"Partial","Non-Compliant")))'
        apply_style(ws[f'F{data_row}'], styles['formula_cell'])
        
        # Columns G-P: Input cells
        for col_letter in ['G', 'H', 'I', 'J', 'L', 'M', 'N', 'O', 'P']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column G, I, J: Datetime format
        for col_letter in ['G', 'I', 'J']:
            ws[f'{col_letter}{data_row}'].number_format = 'DD.MM.YYYY HH:MM'
        
        # Column O: Date format
        ws[f'O{data_row}'].number_format = 'DD.MM.YYYY'
        
        # Column K: Gap Duration Formula
        ws[f'K{data_row}'] = f'=IF(AND(I{data_row}<>"",J{data_row}<>""),(J{data_row}-I{data_row})*24,"")'
        ws[f'K{data_row}'].number_format = '0.0'
        apply_style(ws[f'K{data_row}'], styles['formula_cell'])
    
    # Data validations
    gap_detected_dv = DataValidation(type="list",
        formula1='"Yes,No"',
        allow_blank=True)
    ws.add_data_validation(gap_detected_dv)
    gap_detected_dv.add('H9:H200')
    
    gap_reason_dv = DataValidation(type="list",
        formula1='"Network Issue,Forwarder Down,Source Down,SIEM Issue,Configuration Error,Unknown"',
        allow_blank=True)
    ws.add_data_validation(gap_reason_dv)
    gap_reason_dv.add('L9:L200')
    
    ws.freeze_panes = 'C8'

# ============================================================================
# SECTION 6: INTEGRATION ARCHITECTURE SHEET
# ============================================================================

def create_integration_architecture_sheet(ws, styles):
    """
    Create Integration Architecture sheet.
    
    "Architecture is politics" - Mitch Kapor
    Let's document the political (and technical) reality of our log flows!
    """
    
    # Header
    ws.merge_cells('A1:N1')
    ws['A1'] = "INTEGRATION ARCHITECTURE"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:N2')
    ws['A2'] = "Document log flow architecture and integration points"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        ("A", "Integration Point", 25),
        ("B", "Source Type", 20),
        ("C", "Source Count", 15),
        ("D", "Collection Method", 25),
        ("E", "Intermediate Hops", 20),
        ("F", "Network Path", 30),
        ("G", "Bandwidth Required", 20),
        ("H", "Latency Target", 15),
        ("I", "Current Latency", 15),
        ("J", "Redundancy", 15),
        ("K", "Single Point of Failure", 20),
        ("L", "DR Capability", 15),
        ("M", "Bottleneck Risk", 15),
        ("N", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "Web Servers → SIEM", "Server", "50", "Agent-based", "1",
        "Prod VLAN → Log Concentrator → SIEM VLAN", "100 Mbps", "<1 min",
        "30 sec", "Active/Active", "No", "Yes", "Low", "Direct forwarder deployment"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-50)
    for data_row in range(9, 51):
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
    
    # Data validations
    source_type_dv = DataValidation(type="list",
        formula1='"Server,Network Device,Application,Cloud Service,Security Tool,Container,Other"',
        allow_blank=True)
    ws.add_data_validation(source_type_dv)
    source_type_dv.add('B9:B50')
    
    method_dv = DataValidation(type="list",
        formula1='"Agent-based,Agentless/Syslog,API Pull,API Push,File Collection,Stream Processing"',
        allow_blank=True)
    ws.add_data_validation(method_dv)
    method_dv.add('D9:D50')
    
    redundancy_dv = DataValidation(type="list",
        formula1='"None,Active/Passive,Active/Active"',
        allow_blank=True)
    ws.add_data_validation(redundancy_dv)
    redundancy_dv.add('J9:J50')
    
    spof_dv = DataValidation(type="list",
        formula1='"Yes,No,Mitigated"',
        allow_blank=True)
    ws.add_data_validation(spof_dv)
    spof_dv.add('K9:K50')
    
    dr_dv = DataValidation(type="list",
        formula1='"Yes,No,Partial"',
        allow_blank=True)
    ws.add_data_validation(dr_dv)
    dr_dv.add('L9:L50')
    
    bottleneck_dv = DataValidation(type="list",
        formula1='"High,Medium,Low,None"',
        allow_blank=True)
    ws.add_data_validation(bottleneck_dv)
    bottleneck_dv.add('M9:M50')
    
    # Architecture diagram placeholder
    row = 55
    ws.merge_cells(f'A{row}:N{row}')
    ws[f'A{row}'] = "ARCHITECTURE DIAGRAM"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 1
    ws.merge_cells(f'A{row}:N{row+10}')
    ws[f'A{row}'] = "Insert or link to architecture diagram showing complete log flow:\nSources → Forwarders → SIEM → Storage\n\nInclude network paths, ports, protocols, and redundancy."
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    apply_style(ws[f'A{row}'], styles['input_cell'])
    ws.row_dimensions[row].height = 150
    
    ws.freeze_panes = 'A8'

# ============================================================================
# SECTION 7: SIEM STORAGE & CAPACITY SHEET
# ============================================================================

def create_siem_storage_capacity_sheet(ws, styles):
    """Create SIEM Storage & Capacity sheet."""
    
    # Header
    ws.merge_cells('A1:O1')
    ws['A1'] = "SIEM STORAGE & CAPACITY"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:O2')
    ws['A2'] = "Assess storage capacity, growth trends, and capacity planning"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        ("A", "Storage Component", 25),
        ("B", "Technology", 20),
        ("C", "Total Capacity (TB)", 20),
        ("D", "Used Capacity (TB)", 20),
        ("E", "Free Capacity (TB)", 20),
        ("F", "% Used", 12),
        ("G", "Status", 15),
        ("H", "Retention Period", 20),
        ("I", "Avg Daily Ingest (GB)", 20),
        ("J", "Growth Rate %/Month", 20),
        ("K", "Days Until Full", 15),
        ("L", "Expansion Plan", 30),
        ("M", "Expansion Date", 15),
        ("N", "Cost per TB/Month", 20),
        ("O", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "Hot Storage", "Local SSD", "50", "32", "18", "64%", "OK",
        "90 days", "350", "5", "51", "Not Needed", "", "$15", "Primary indexer storage"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-20)
    for data_row in range(9, 21):
        # Columns A-B, H-J, L, N-O: Input cells
        for col_letter in ['A', 'B', 'H', 'I', 'J', 'L', 'N', 'O']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Columns C-D: Input (capacity numbers)
        for col_letter in ['C', 'D']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column E: Free Capacity Formula
        ws[f'E{data_row}'] = f'=IF(C{data_row}="","",C{data_row}-D{data_row})'
        apply_style(ws[f'E{data_row}'], styles['formula_cell'])
        
        # Column F: % Used Formula
        ws[f'F{data_row}'] = f'=IF(C{data_row}=0,0,D{data_row}/C{data_row})'
        ws[f'F{data_row}'].number_format = '0.0%'
        apply_style(ws[f'F{data_row}'], styles['formula_cell'])
        
        # Column G: Status Formula
        ws[f'G{data_row}'] = f'=IF(A{data_row}="","",IF(F{data_row}>0.85,"Critical",IF(F{data_row}>0.7,"Warning","OK")))'
        apply_style(ws[f'G{data_row}'], styles['formula_cell'])
        
        # Column K: Days Until Full Formula
        ws[f'K{data_row}'] = f'=IF(OR(I{data_row}=0,E{data_row}=""),"N/A",(E{data_row}*1024)/I{data_row})'
        ws[f'K{data_row}'].number_format = '0'
        apply_style(ws[f'K{data_row}'], styles['formula_cell'])
        
        # Column M: Expansion Date
        apply_style(ws[f'M{data_row}'], styles['input_cell'])
        ws[f'M{data_row}'].number_format = 'DD.MM.YYYY'
    
    # Data validations
    storage_component_dv = DataValidation(type="list",
        formula1='"Hot Storage,Warm Storage,Cold Storage,Archive,Backup"',
        allow_blank=True)
    ws.add_data_validation(storage_component_dv)
    storage_component_dv.add('A9:A20')
    
    technology_dv = DataValidation(type="list",
        formula1='"Local Disk/SSD,SAN,NAS,Object Storage (S3/Azure),Tape,Cloud,Other"',
        allow_blank=True)
    ws.add_data_validation(technology_dv)
    technology_dv.add('B9:B20')
    
    expansion_dv = DataValidation(type="list",
        formula1='"Not Needed,Planned,In Progress,Urgent"',
        allow_blank=True)
    ws.add_data_validation(expansion_dv)
    expansion_dv.add('L9:L20')
    
    # Capacity Planning Section
    row = 25
    ws.merge_cells(f'A{row}:O{row}')
    ws[f'A{row}'] = "CAPACITY PLANNING"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    # Planning table headers
    planning_headers = ["Timeframe", "Projected Ingest (TB)", "Capacity Needed (TB)", "Gap (TB)", "Estimated Cost"]
    for col_idx, header in enumerate(planning_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row += 1
    timeframes = [
        ("Current", "=SUM(D9:D20)", "=SUM(C9:C20)", "N/A", "N/A"),
        ("6 months", "=SUM(I9:I20)*180/1024", "[Formula]", "[Formula]", "[Input]"),
        ("12 months", "=SUM(I9:I20)*365/1024", "[Formula]", "[Formula]", "[Input]"),
        ("24 months", "=SUM(I9:I20)*730/1024", "[Formula]", "[Formula]", "[Input]"),
    ]
    
    for timeframe, proj_ingest, capacity_needed, gap, cost in timeframes:
        ws[f'A{row}'] = timeframe
        ws[f'A{row}'].font = Font(bold=True, size=10)
        
        ws[f'B{row}'] = proj_ingest
        ws[f'C{row}'] = capacity_needed
        ws[f'D{row}'] = gap
        ws[f'E{row}'] = cost
        
        for col in ['B', 'C', 'D', 'E']:
            if "[" in str(ws.cell(row=row, column=ord(col)-64).value):
                apply_style(ws.cell(row=row, column=ord(col)-64), styles['input_cell'])
            else:
                apply_style(ws.cell(row=row, column=ord(col)-64), styles['formula_cell'])
        
        row += 1
    
    # Optimization opportunities
    row += 2
    ws.merge_cells(f'A{row}:O{row}')
    ws[f'A{row}'] = "OPTIMIZATION OPPORTUNITIES"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 1
    opportunities = [
        "\u2022 Enable compression to reduce storage usage by 50-70%",
        "\u2022 Implement tiered storage (hot/warm/cold) for cost optimization",
        "\u2022 Review retention periods - are all logs kept longer than necessary?",
        "\u2022 Archive old data to low-cost object storage (S3 Glacier, Azure Archive)",
        "\u2022 Implement data summarization for aged logs",
    ]
    
    for opportunity in opportunities:
        ws.merge_cells(f'A{row}:O{row}')
        ws[f'A{row}'] = opportunity
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        row += 1
    
    ws.freeze_panes = 'A8'

# ============================================================================
# SECTION 8: LOG PARSING & NORMALIZATION SHEET
# ============================================================================

def create_log_parsing_normalization_sheet(ws, styles):
    """Create Log Parsing & Normalization sheet."""
    
    # Header
    ws.merge_cells('A1:O1')
    ws['A1'] = "LOG PARSING & NORMALIZATION"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:O2')
    ws['A2'] = "Assess parsing accuracy, field extraction, and data normalization"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        ("A", "Log Source Type", 25),
        ("B", "Log Format", 20),
        ("C", "Parsing Method", 20),
        ("D", "Parser Status", 15),
        ("E", "Parse Success Rate %", 20),
        ("F", "Unparsed Events/Day", 20),
        ("G", "Fields Extracted", 20),
        ("H", "Required Fields Present", 20),
        ("I", "Timestamp Parsing", 18),
        ("J", "Timezone Handling", 18),
        ("K", "Last Parser Update", 15),
        ("L", "Issues Identified", 40),
        ("M", "Tuning Required", 15),
        ("N", "Owner", 20),
        ("O", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "Windows Events", "EVTX", "Built-in", "Working", "98.5", "3750",
        "25", "All", "Correct", "Correct", "15.11.2025", "None", "No",
        "SIEM Team", "Standard Windows parser"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-100)
    for data_row in range(9, 101):
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'M', 'N', 'O']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column K: Date format
        ws[f'K{data_row}'].number_format = 'DD.MM.YYYY'
        
        # Column E: Format as percentage
        ws[f'E{data_row}'].number_format = '0.0'
    
    # Data validations
    log_format_dv = DataValidation(type="list",
        formula1='"Syslog,CEF,JSON,EVTX,LEEF,Custom"',
        allow_blank=True)
    ws.add_data_validation(log_format_dv)
    log_format_dv.add('B9:B100')
    
    parsing_method_dv = DataValidation(type="list",
        formula1='"Built-in,Custom Regex,Grok,Logstash Filter,Custom Script,ML-based"',
        allow_blank=True)
    ws.add_data_validation(parsing_method_dv)
    parsing_method_dv.add('C9:C100')
    
    parser_status_dv = DataValidation(type="list",
        formula1='"Working,Needs Tuning,Broken,Not Configured"',
        allow_blank=True)
    ws.add_data_validation(parser_status_dv)
    parser_status_dv.add('D9:D100')
    
    required_fields_dv = DataValidation(type="list",
        formula1='"All,Most,Some,Few,None"',
        allow_blank=True)
    ws.add_data_validation(required_fields_dv)
    required_fields_dv.add('H9:H100')
    
    timestamp_dv = DataValidation(type="list",
        formula1='"Correct,Incorrect,Missing"',
        allow_blank=True)
    ws.add_data_validation(timestamp_dv)
    timestamp_dv.add('I9:I100')
    
    timezone_dv = DataValidation(type="list",
        formula1='"Correct,Incorrect,Unknown"',
        allow_blank=True)
    ws.add_data_validation(timezone_dv)
    timezone_dv.add('J9:J100')
    
    tuning_dv = DataValidation(type="list",
        formula1='"Yes,No"',
        allow_blank=True)
    ws.add_data_validation(tuning_dv)
    tuning_dv.add('M9:M100')
    
    ws.freeze_panes = 'A8'

# ============================================================================
# SECTION 9: SIEM PERFORMANCE METRICS SHEET
# ============================================================================

def create_siem_performance_metrics_sheet(ws, styles):
    """Create SIEM Performance Metrics sheet."""
    
    # Header
    ws.merge_cells('A1:N1')
    ws['A1'] = "SIEM PERFORMANCE METRICS"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:N2')
    ws['A2'] = "Track SIEM platform performance over assessment period (daily metrics)"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        ("A", "Metric Date", 15),
        ("B", "Daily Events Indexed", 20),
        ("C", "Peak Events/Second", 20),
        ("D", "Indexing Lag (minutes)", 20),
        ("E", "Search Performance (sec)", 20),
        ("F", "Dashboard Load Time (sec)", 20),
        ("G", "CPU Utilization %", 18),
        ("H", "Memory Utilization %", 18),
        ("I", "Disk I/O %", 15),
        ("J", "Network Throughput (Gbps)", 20),
        ("K", "Uptime %", 12),
        ("L", "Incidents", 15),
        ("M", "Performance Status", 18),
        ("N", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "06.01.2026", "25000000", "8500", "3", "8", "5",
        "65", "72", "45", "2.5", "100", "0", "Healthy", "Normal operations"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-90 for 30-90 days of daily metrics)
    for data_row in range(9, 91):
        # Column A: Date
        apply_style(ws[f'A{data_row}'], styles['input_cell'])
        ws[f'A{data_row}'].number_format = 'DD.MM.YYYY'
        
        # Columns B-L: Input cells (metrics)
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column M: Performance Status Formula
        # Healthy if: Indexing lag <=5, Search <=10, CPU/Mem <80%, Uptime >=99.9%
        ws[f'M{data_row}'] = f'=IF(A{data_row}="","",IF(AND(D{data_row}<=5,E{data_row}<=10,G{data_row}<80,H{data_row}<80,K{data_row}>=99.9),"Healthy",IF(OR(D{data_row}>15,E{data_row}>30,G{data_row}>90,H{data_row}>90,K{data_row}<99),"Critical","Warning")))'
        apply_style(ws[f'M{data_row}'], styles['formula_cell'])
        
        # Column N: Notes
        apply_style(ws[f'N{data_row}'], styles['input_cell'])
    
    # Summary section
    row = 95
    ws.merge_cells(f'A{row}:N{row}')
    ws[f'A{row}'] = "PERFORMANCE SUMMARY (Last 30 Days)"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    summary_metrics = [
        ("Avg Daily Events Indexed:", "=AVERAGE(B9:B38)"),
        ("Avg Peak EPS:", "=AVERAGE(C9:C38)"),
        ("Avg Indexing Lag (min):", "=AVERAGE(D9:D38)"),
        ("Avg Search Performance (sec):", "=AVERAGE(E9:E38)"),
        ("Avg CPU Utilization %:", "=AVERAGE(G9:G38)"),
        ("Avg Memory Utilization %:", "=AVERAGE(H9:H38)"),
        ("Avg Uptime %:", "=AVERAGE(K9:K38)"),
        ("Total Incidents:", "=SUM(L9:L38)"),
        ("Days with Healthy Status:", "=COUNTIF(M9:M38,\"Healthy\")"),
    ]
    
    for label, formula in summary_metrics:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = formula
        apply_style(ws[f'B{row}'], styles['formula_cell'])
        row += 1
    
    ws.freeze_panes = 'A8'

# ============================================================================
# SECTION 10: DATA QUALITY ASSESSMENT SHEET
# ============================================================================

def create_data_quality_assessment_sheet(ws, styles):
    """Create Data Quality Assessment sheet."""
    
    # Header
    ws.merge_cells('A1:N1')
    ws['A1'] = "DATA QUALITY ASSESSMENT"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:N2')
    ws['A2'] = "Assess quality of log data in SIEM (completeness, accuracy, consistency, timeliness, validity)"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        ("A", "Quality Dimension", 25),
        ("B", "Log Source Type", 25),
        ("C", "Assessment Method", 30),
        ("D", "Sample Size", 15),
        ("E", "Pass Count", 15),
        ("F", "Fail Count", 15),
        ("G", "Quality Score %", 18),
        ("H", "Status", 15),
        ("I", "Common Issues", 40),
        ("J", "Impact", 20),
        ("K", "Remediation Action", 40),
        ("L", "Responsible Party", 25),
        ("M", "Target Date", 15),
        ("N", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "Completeness", "Web Server Logs", "Field presence check on 1000 events",
        "1000", "985", "15", "98.5%", "Good", "Missing source_ip in 15 events",
        "Low", "Update parser to extract source_ip", "SIEM Team", "31.01.2026",
        "Minor issue, easily fixable"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-50)
    for data_row in range(9, 51):
        # Columns A-F, I-L, N: Input cells
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'I', 'J', 'K', 'L', 'N']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column G: Quality Score Formula
        ws[f'G{data_row}'] = f'=IF(AND(D{data_row}<>"",E{data_row}<>""),E{data_row}/(E{data_row}+F{data_row}),"")'
        ws[f'G{data_row}'].number_format = '0.0%'
        apply_style(ws[f'G{data_row}'], styles['formula_cell'])
        
        # Column H: Status Formula
        ws[f'H{data_row}'] = f'=IF(G{data_row}="","",IF(G{data_row}>=0.95,"Good",IF(G{data_row}>=0.8,"Fair","Poor")))'
        apply_style(ws[f'H{data_row}'], styles['formula_cell'])
        
        # Column M: Target Date
        apply_style(ws[f'M{data_row}'], styles['input_cell'])
        ws[f'M{data_row}'].number_format = 'DD.MM.YYYY'
    
    # Data validations
    quality_dimension_dv = DataValidation(type="list",
        formula1='"Completeness,Accuracy,Consistency,Timeliness,Validity"',
        allow_blank=True)
    ws.add_data_validation(quality_dimension_dv)
    quality_dimension_dv.add('A9:A50')
    
    impact_dv = DataValidation(type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚫ Low"',
        allow_blank=True)
    ws.add_data_validation(impact_dv)
    impact_dv.add('J9:J50')
    
    # Quality dimensions explanation
    row = 55
    ws.merge_cells(f'A{row}:N{row}')
    ws[f'A{row}'] = "QUALITY DIMENSIONS EXPLAINED"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    explanations = [
        ("Completeness:", "All required fields present in logs"),
        ("Accuracy:", "Data values correct and meaningful"),
        ("Consistency:", "Same events logged same way across sources"),
        ("Timeliness:", "Logs indexed within target timeframe"),
        ("Validity:", "Data conforms to expected formats"),
    ]
    
    for dimension, explanation in explanations:
        ws[f'A{row}'] = dimension
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = explanation
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(f'B{row}:N{row}')
        row += 1
    
    ws.freeze_panes = 'A8'


# ============================================================================
# SECTION 10.5: ENCRYPTION & AUTHENTICATION SHEET
# ============================================================================

def create_encryption_authentication_sheet(ws, styles):
    """Create Encryption & Authentication sheet for log transport security."""

    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "ENCRYPTION & AUTHENTICATION"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:K2')
    ws['A2'] = "Assess security controls for log transport and collection authentication"
    apply_style(ws['A2'], styles['header_sub'])

    # ==================== TLS/ENCRYPTION ASSESSMENT ====================
    row = 4
    ws.merge_cells(f'A{row}:K{row}')
    ws[f'A{row}'] = "LOG TRANSPORT ENCRYPTION"
    apply_style(ws[f'A{row}'], styles['section_header'])
    row += 1

    headers = ["Log Source/Path", "Protocol", "TLS Version", "Cipher Suite", "Certificate Valid", "Certificate Expiry", "Mutual TLS", "Compliance Status", "Last Verified", "Evidence Ref", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles['column_header'])
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['K'].width = 30
    row += 1

    # Sample log paths
    log_paths = [
        ("Syslog -> SIEM", "Syslog/TLS", "", "", "", "", "", "", "", "", ""),
        ("Windows Event -> SIEM", "WEF/HTTPS", "", "", "", "", "", "", "", "", ""),
        ("Linux Agents -> SIEM", "Agent/TLS", "", "", "", "", "", "", "", "", ""),
        ("Cloud Logs -> SIEM", "HTTPS API", "", "", "", "", "", "", "", "", ""),
        ("Network Devices -> SIEM", "Syslog/TLS", "", "", "", "", "", "", "", "", ""),
        ("Database Audit -> SIEM", "JDBC/TLS", "", "", "", "", "", "", "", "", ""),
        ("Container Logs -> SIEM", "Fluentd/TLS", "", "", "", "", "", "", "", "", ""),
        ("Application Logs -> SIEM", "Filebeat/TLS", "", "", "", "", "", "", "", "", ""),
    ]

    for path_data in log_paths:
        for col_idx, value in enumerate(path_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 2:
                apply_style(cell, styles['input_cell'])
        row += 1

    row += 2

    # ==================== AUTHENTICATION METHODS ====================
    ws.merge_cells(f'A{row}:K{row}')
    ws[f'A{row}'] = "AUTHENTICATION METHODS"
    apply_style(ws[f'A{row}'], styles['section_header'])
    row += 1

    auth_headers = ["Collection Method", "Auth Type", "Credential Storage", "Rotation Frequency", "Last Rotated", "Service Account", "MFA Enabled", "Compliance", "Owner", "Evidence Ref"]
    for col_idx, header in enumerate(auth_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles['column_header'])
    row += 1

    auth_methods = [
        ("SIEM API Access", "API Key", "", "", "", "", "", "", "", ""),
        ("Log Forwarder Auth", "Certificate", "", "", "", "", "", "", "", ""),
        ("Cloud Platform Auth", "OAuth/OIDC", "", "", "", "", "", "", "", ""),
        ("Database Audit Auth", "Service Account", "", "", "", "", "", "", "", ""),
        ("Agent Authentication", "Certificate", "", "", "", "", "", "", "", ""),
    ]

    for auth_data in auth_methods:
        for col_idx, value in enumerate(auth_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 2:
                apply_style(cell, styles['input_cell'])
        row += 1

    row += 2

    # ==================== COMPLIANCE CHECKLIST ====================
    ws.merge_cells(f'A{row}:K{row}')
    ws[f'A{row}'] = "ENCRYPTION & AUTHENTICATION COMPLIANCE CHECKLIST"
    apply_style(ws[f'A{row}'], styles['section_header'])
    row += 1

    checklist_items = [
        ("All log transport uses TLS 1.2 or higher", ""),
        ("No plaintext log transmission over network", ""),
        ("Certificates are valid and not expired", ""),
        ("Certificate expiry monitoring in place", ""),
        ("Mutual TLS (mTLS) used where supported", ""),
        ("API keys/tokens stored securely (vault/KMS)", ""),
        ("Credentials rotated per policy (90 days max)", ""),
        ("Service accounts have minimum required permissions", ""),
        ("Authentication failures are logged and alerted", ""),
        ("Encryption at rest enabled for log storage", ""),
    ]

    for item, status in checklist_items:
        ws[f'A{row}'] = item
        ws[f'A{row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws.cell(row=row, column=10, value=status)
        apply_style(cell, styles['input_cell'])
        row += 1

    ws.freeze_panes = 'A6'


# ============================================================================
# SECTION 11: GAP ANALYSIS & REMEDIATION SHEET
# ============================================================================

def create_gap_analysis_sheet(ws, styles):
    """Create Gap Analysis & Remediation sheet."""
    
    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "GAP ANALYSIS & REMEDIATION"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:L2')
    ws['A2'] = "Track collection infrastructure gaps and remediation plans"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 30
    
    # Column headers
    headers = [
        ("A", "Gap ID", 12),
        ("B", "Gap Category", 25),
        ("C", "Description", 50),
        ("D", "Affected Systems", 30),
        ("E", "Impact", 20),
        ("F", "Current State", 40),
        ("G", "Target State", 40),
        ("H", "Remediation Action", 50),
        ("I", "Owner", 25),
        ("J", "Target Date", 15),
        ("K", "Status", 15),
        ("L", "Notes", 40),
    ]
    
    row = 7
    for col_letter, header, width in headers:
        ws[f'{col_letter}{row}'] = header
        apply_style(ws[f'{col_letter}{row}'], styles['column_header'])
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[row].height = 35
    
    # Example row
    row = 8
    example_data = [
        "GAP-001", "Reliability Issue", "Collection rate below 95% for 15 systems",
        "15 database servers", "High", "Avg collection rate: 87%",
        "Collection rate: >95%", "Investigate forwarder configuration and network path",
        "DB Admin Team", "31.01.2026", "Open", "May require forwarder upgrade"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        apply_style(cell, styles['example_cell'])
    
    # Data entry rows (9-100)
    for data_row in range(9, 101):
        # Column A: Auto-generate Gap ID
        ws[f'A{data_row}'] = f'=IF(B{data_row}<>"","GAP-"&TEXT(ROW()-8,"000"),"")'
        apply_style(ws[f'A{data_row}'], styles['formula_cell'])
        
        # Columns B-I, K-L: Input cells
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'K', 'L']:
            apply_style(ws[f'{col_letter}{data_row}'], styles['input_cell'])
        
        # Column J: Target Date
        apply_style(ws[f'J{data_row}'], styles['input_cell'])
        ws[f'J{data_row}'].number_format = 'DD.MM.YYYY'
    
    # Data validations
    gap_category_dv = DataValidation(type="list",
        formula1='"Coverage Gap,Reliability Issue,Performance Issue,Parsing Error,Capacity Constraint,Redundancy Gap,DR Gap,Data Quality Issue"',
        allow_blank=True)
    ws.add_data_validation(gap_category_dv)
    gap_category_dv.add('B9:B100')
    
    impact_dv = DataValidation(type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚫ Low"',
        allow_blank=True)
    ws.add_data_validation(impact_dv)
    impact_dv.add('E9:E100')
    
    status_dv = DataValidation(type="list",
        formula1='"\u274C Open,⏳ In Progress,\u2705 Resolved,⭕ Deferred"',
        allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add('K9:K100')
    
    ws.freeze_panes = 'A8'


# ============================================================================
# SECTION 11.5: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register_sheet(ws, styles):
    """Create Evidence Register for audit documentation."""

    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "EVIDENCE REGISTER"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:L2')
    ws['A2'] = "Document all supporting evidence for log collection assessment"
    apply_style(ws['A2'], styles['header_sub'])

    ws.merge_cells('A3:L3')
    ws['A3'] = "Reference evidence by ID in other assessment sheets. Store evidence files in designated secure location."
    ws['A3'].font = Font(italic=True, size=9)
    ws['A3'].alignment = Alignment(horizontal='left')

    # Evidence table headers
    row = 5
    headers = [
        ("Evidence ID", 12),
        ("Assessment Sheet", 22),
        ("Evidence Type", 20),
        ("Evidence Title", 35),
        ("Description", 40),
        ("File Location/Link", 35),
        ("Date Collected", 14),
        ("Collected By", 18),
        ("Retention Period", 15),
        ("Review Date", 14),
        ("Status", 12),
        ("Notes", 30),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles['column_header'])
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    row += 1

    # Pre-populate with common evidence types
    evidence_items = [
        ("EVD-001", "SIEM Platform Details", "Configuration", "SIEM Architecture Diagram", "", "", "", "", "7 years", "", "", ""),
        ("EVD-002", "SIEM Platform Details", "Screenshot", "SIEM License/Version", "", "", "", "", "7 years", "", "", ""),
        ("EVD-003", "Log Forwarder Inventory", "Export", "Agent Inventory Export", "", "", "", "", "7 years", "", "", ""),
        ("EVD-004", "Collection Reliability", "Report", "Collection Success Rate Report", "", "", "", "", "7 years", "", "", ""),
        ("EVD-005", "Encryption & Authentication", "Certificate", "TLS Certificate Details", "", "", "", "", "7 years", "", "", ""),
        ("EVD-006", "Encryption & Authentication", "Configuration", "Authentication Config Export", "", "", "", "", "7 years", "", "", ""),
        ("EVD-007", "SIEM Storage & Capacity", "Report", "Storage Utilization Report", "", "", "", "", "7 years", "", "", ""),
        ("EVD-008", "Gap Analysis & Remediation", "Document", "Gap Remediation Plan", "", "", "", "", "7 years", "", "", ""),
    ]

    for evidence_data in evidence_items:
        for col_idx, value in enumerate(evidence_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 4:
                apply_style(cell, styles['input_cell'])
        row += 1

    # Add 50 more empty rows for evidence
    for i in range(50):
        evd_num = f"EVD-{9+i:03d}"
        ws.cell(row=row, column=1, value=evd_num)
        for col_idx in range(2, 13):
            cell = ws.cell(row=row, column=col_idx, value="")
            apply_style(cell, styles['input_cell'])
        row += 1

    # Data validations
    type_dv = DataValidation(type="list",
        formula1='"Configuration,Screenshot,Export,Report,Certificate,Document,Log Sample,Test Result,Approval,Other"',
        allow_blank=True)
    ws.add_data_validation(type_dv)
    type_dv.add('C6:C100')

    sheet_dv = DataValidation(type="list",
        formula1='"SIEM Platform Details,Log Forwarder Inventory,Collection Reliability,Integration Architecture,SIEM Storage & Capacity,Log Parsing & Normalization,SIEM Performance Metrics,Data Quality Assessment,Encryption & Authentication,Gap Analysis & Remediation"',
        allow_blank=True)
    ws.add_data_validation(sheet_dv)
    sheet_dv.add('B6:B100')

    status_dv = DataValidation(type="list",
        formula1='"Collected,Pending,Expired,Not Available"',
        allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add('K6:K100')

    ws.freeze_panes = 'A6'


# ============================================================================
# SECTION 12: SUMMARY DASHBOARD SHEET
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """
    Create Summary Dashboard - executive view.

    "Simplicity is the ultimate sophistication." - Leonardo da Vinci
    Let's make this dashboard simple yet sophisticated!
    """
    
    # Main header
    ws.merge_cells('A1:H1')
    ws['A1'] = "LOG COLLECTION & CENTRALIZATION DASHBOARD"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:H2')
    ws['A2'] = "Executive Summary - SIEM Health & Performance"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 25
    
    # Section 1: Collection Health Summary
    row = 4
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "1. COLLECTION HEALTH SUMMARY"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    metrics = [
        ("Total Systems Monitored", "=COUNTA('Collection Reliability'!A9:A200)", "N/A"),
        ("Systems Collecting >95%", "=COUNTIF('Collection Reliability'!E:E,\">0.95\")", "100%"),
        ("Avg Collection Rate %", "=AVERAGE('Collection Reliability'!E9:E200)", ">98%"),
        ("Systems with Gaps >4hr", "=COUNTIF('Collection Reliability'!K:K,\">4\")", "0"),
        ("Parse Success Rate %", "=AVERAGE('Log Parsing & Normalization'!E9:E100)", ">95%"),
        ("SIEM Uptime %", "=AVERAGE('SIEM Performance Metrics'!K9:K38)", ">99.9%"),
        ("Storage Used %", "=AVERAGE('SIEM Storage & Capacity'!F9:F20)", "<70%"),
        ("Days Until Storage Full", "=MIN('SIEM Storage & Capacity'!K9:K20)", ">90"),
    ]
    
    # Headers
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Target"
    ws[f'D{row}'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    row += 1
    for metric_name, formula, target in metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(bold=True, size=10)
        
        ws[f'B{row}'] = formula
        apply_style(ws[f'B{row}'], styles['formula_cell'])
        if "Rate" in metric_name or "Uptime" in metric_name or "Used" in metric_name:
            ws[f'B{row}'].number_format = '0.0%'
        
        ws[f'C{row}'] = target
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Status column
        ws[f'D{row}'] = "=IF(B{row}>0,\"✓\",\"✗\")"
        apply_style(ws[f'D{row}'], styles['formula_cell'])
        
        row += 1
    
    # Section 2: Collection by System Type
    row += 2
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "2. COLLECTION BY SYSTEM TYPE"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    ws[f'A{row}'] = "System Type"
    ws[f'B{row}'] = "Total"
    ws[f'C{row}'] = "Collecting >95%"
    ws[f'D{row}'] = "Avg Collection Rate %"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], styles['column_header'])
    
    row += 1
    system_types = ["Server", "Network Device", "Security Appliance", "Application", "Cloud Service"]
    for sys_type in system_types:
        ws[f'A{row}'] = sys_type
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = f'=COUNTIF(\'Collection Reliability\'!B:B,"*{sys_type}*")'
        ws[f'C{row}'] = f'=COUNTIFS(\'Collection Reliability\'!B:B,"*{sys_type}*",\'Collection Reliability\'!E:E,">0.95")'
        ws[f'D{row}'] = f'=AVERAGEIF(\'Collection Reliability\'!B:B,"*{sys_type}*",\'Collection Reliability\'!E:E)'
        ws[f'D{row}'].number_format = '0.0%'
        
        for col in ['B', 'C', 'D']:
            apply_style(ws[f'{col}{row}'], styles['formula_cell'])
        
        row += 1
    
    # Section 3: Gap Summary
    row += 2
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "3. GAP SUMMARY BY PRIORITY"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    gap_headers = ["Priority", "Open", "In Progress", "Resolved", "Overdue"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row += 1
    priorities = ["Critical", "High", "Medium", "Low"]
    
    for priority in priorities:
        ws[f'A{row}'] = priority
        ws[f'A{row}'].font = Font(bold=True, size=10)
        
        ws[f'B{row}'] = f'=COUNTIFS(\'Gap Analysis & Remediation\'!E:E,"{priority}",\'Gap Analysis & Remediation\'!K:K,"Open")'
        ws[f'C{row}'] = f'=COUNTIFS(\'Gap Analysis & Remediation\'!E:E,"{priority}",\'Gap Analysis & Remediation\'!K:K,"In Progress")'
        ws[f'D{row}'] = f'=COUNTIFS(\'Gap Analysis & Remediation\'!E:E,"{priority}",\'Gap Analysis & Remediation\'!K:K,"Resolved")'
        ws[f'E{row}'] = f'=SUMPRODUCT((\'Gap Analysis & Remediation\'!E:E="{priority}")*((\'Gap Analysis & Remediation\'!K:K="Open")+(\'Gap Analysis & Remediation\'!K:K="In Progress"))*(\'Gap Analysis & Remediation\'!J:J<TODAY())*(\'Gap Analysis & Remediation\'!J:J<>""))'
        
        for col in ['B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], styles['formula_cell'])
        
        row += 1
    
    # Section 4: Recommendations
    row += 2
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "4. RECOMMENDATIONS"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    recommendations = [
        "1. If avg collection rate <98%: Investigate and resolve collection gaps",
        "2. If storage >70%: Plan storage expansion within 90 days",
        "3. If parse success rate <95%: Tune parsers for improved field extraction",
        "4. If SIEM uptime <99.9%: Implement redundancy/high availability",
        "5. If critical gaps exist: Prioritize remediation within 30 days",
        "6. Review forwarder versions: Ensure all forwarders are up-to-date",
        "7. Implement automated monitoring: Set up alerts for collection failures",
        "8. Document runbooks: Create procedures for common issues",
    ]
    
    for recommendation in recommendations:
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = recommendation
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if recommendation[0].isdigit():
            ws[f'A{row}'].font = Font(size=9)
        row += 1
    
    # Set column widths
    set_column_widths(ws, {
        'A': 35, 'B': 20, 'C': 15, 'D': 15,
        'E': 20, 'F': 20, 'G': 20, 'H': 20
    })

# ============================================================================
# SECTION 13: APPROVAL & SIGN-OFF SHEET
# ============================================================================

def create_approval_signoff_sheet(ws, styles):
    """Create Approval & Sign-Off sheet."""
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "APPROVAL & SIGN-OFF"
    apply_style(ws['A1'], styles['header_main'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:E2')
    ws['A2'] = "Multi-level approval workflow for Log Collection & Centralization Assessment"
    apply_style(ws['A2'], styles['header_sub'])
    ws.row_dimensions[2].height = 25
    
    # Approval table
    row = 4
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "APPROVAL WORKFLOW"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    approval_headers = ["Role", "Name", "Date", "Signature", "Status"]
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    row += 1
    approval_roles = [
        ("Log Administrator", "[Name]", "", "_____", "☐ Reviewed"),
        ("SOC Lead", "[Name]", "", "_____", "☐ Reviewed"),
        ("IT Operations Manager", "[Name]", "", "_____", "☐ Reviewed"),
        ("Information Security Manager", "[Name]", "", "_____", "☐ Approved"),
    ]
    
    for role, name, date, signature, status in approval_roles:
        ws[f'A{row}'] = role
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = name
        apply_style(ws[f'B{row}'], styles['input_cell'])
        ws[f'C{row}'] = date
        apply_style(ws[f'C{row}'], styles['input_cell'])
        ws[f'C{row}'].number_format = 'DD.MM.YYYY'
        ws[f'D{row}'] = signature
        apply_style(ws[f'D{row}'], styles['input_cell'])
        ws[f'E{row}'] = status
        ws[f'E{row}'].alignment = Alignment(horizontal='left', vertical='center')
        row += 1
    
    # Acknowledgments
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "ACKNOWLEDGMENTS CHECKLIST"
    apply_style(ws[f'A{row}'], styles['header_sub'])
    
    row += 2
    acknowledgments = [
        "☐ SIEM platform architecture documented",
        "☐ Log forwarders inventoried",
        "☐ Collection reliability assessed",
        "☐ Performance metrics tracked",
        "☐ Storage capacity reviewed",
        "☐ Parsing accuracy assessed",
        "☐ Data quality evaluated",
        "☐ Gaps identified and prioritized",
        "☐ Remediation plan established",
        "☐ Next review scheduled",
    ]
    
    for acknowledgment in acknowledgments:
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = acknowledgment
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        row += 1
    
    set_column_widths(ws, {'A': 35, 'B': 25, 'C': 15, 'D': 15, 'E': 20})

# ============================================================================
# SECTION 14: CONDITIONAL FORMATTING
# ============================================================================

def apply_conditional_formatting(wb):
    """Apply conditional formatting across sheets."""
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    # Collection Reliability sheet
    ws = wb['Collection Reliability']
    ws.conditional_formatting.add('E9:E200',
        CellIsRule(operator='greaterThanOrEqual', formula=['0.95'], fill=green_fill))
    ws.conditional_formatting.add('E9:E200',
        CellIsRule(operator='between', formula=['0.8', '0.94'], fill=yellow_fill))
    ws.conditional_formatting.add('E9:E200',
        CellIsRule(operator='lessThan', formula=['0.8'], fill=red_fill))
    
    # SIEM Storage & Capacity sheet
    ws = wb['SIEM Storage & Capacity']
    ws.conditional_formatting.add('F9:F20',
        CellIsRule(operator='lessThan', formula=['0.7'], fill=green_fill))
    ws.conditional_formatting.add('F9:F20',
        CellIsRule(operator='between', formula=['0.7', '0.85'], fill=yellow_fill))
    ws.conditional_formatting.add('F9:F20',
        CellIsRule(operator='greaterThan', formula=['0.85'], fill=red_fill))
    
    # Gap Analysis sheet
    ws = wb['Gap Analysis & Remediation']
    ws.conditional_formatting.add('E9:E100',
        CellIsRule(operator='equal', formula=['"Critical"'], fill=red_fill))
    ws.conditional_formatting.add('E9:E100',
        CellIsRule(operator='equal', formula=['"High"'], fill=orange_fill))
    ws.conditional_formatting.add('E9:E100',
        CellIsRule(operator='equal', formula=['"Medium"'], fill=yellow_fill))
    ws.conditional_formatting.add('E9:E100',
        CellIsRule(operator='equal', formula=['"Low"'], fill=green_fill))
    
    ws.conditional_formatting.add('K9:K100',
        CellIsRule(operator='equal', formula=['"Open"'], fill=red_fill))
    ws.conditional_formatting.add('K9:K100',
        CellIsRule(operator='equal', formula=['"In Progress"'], fill=yellow_fill))
    ws.conditional_formatting.add('K9:K100',
        CellIsRule(operator='equal', formula=['"Resolved"'], fill=green_fill))
    
# ============================================================================
# SECTION 15: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """Main execution function."""
    
    logger.info("=" * 78)
    logger.info("ISMS-IMP-A.8.15.2 - Log Collection & Centralization Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.8.15: Logging")
    logger.info("=" * 78)
    logger.info("")
    
    wb = create_workbook()
    styles = setup_styles()
    
    logger.info("[1/14] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)

    logger.info("[2/14] Creating SIEM Platform Details...")
    create_siem_platform_sheet(wb["SIEM Platform Details"], styles)

    logger.info("[3/14] Creating Log Forwarder Inventory...")
    create_log_forwarder_inventory_sheet(wb["Log Forwarder Inventory"], styles)

    logger.info("[4/14] Creating Collection Reliability...")
    create_collection_reliability_sheet(wb["Collection Reliability"], styles)

    logger.info("[5/14] Creating Integration Architecture...")
    create_integration_architecture_sheet(wb["Integration Architecture"], styles)

    logger.info("[6/14] Creating SIEM Storage & Capacity...")
    create_siem_storage_capacity_sheet(wb["SIEM Storage & Capacity"], styles)

    logger.info("[7/14] Creating Log Parsing & Normalization...")
    create_log_parsing_normalization_sheet(wb["Log Parsing & Normalization"], styles)

    logger.info("[8/14] Creating SIEM Performance Metrics...")
    create_siem_performance_metrics_sheet(wb["SIEM Performance Metrics"], styles)
    
    logger.info("[9/14] Creating Data Quality Assessment...")
    create_data_quality_assessment_sheet(wb["Data Quality Assessment"], styles)

    logger.info("[10/14] Creating Encryption & Authentication...")
    create_encryption_authentication_sheet(wb["Encryption & Authentication"], styles)

    logger.info("[11/14] Creating Gap Analysis & Remediation...")
    create_gap_analysis_sheet(wb["Gap Analysis & Remediation"], styles)

    logger.info("[12/14] Creating Evidence Register...")
    create_evidence_register_sheet(wb["Evidence Register"], styles)

    logger.info("[13/14] Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

    logger.info("[14/14] Creating Approval & Sign-Off...")
    create_approval_signoff_sheet(wb["Approval & Sign-Off"], styles)
    
    logger.info("")
    logger.info("Applying conditional formatting...")
    apply_conditional_formatting(wb)
    
    filename = f"ISMS-IMP-A.8.15.2_Log_Collection_Centralization_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    logger.info("")
    logger.info("Saving workbook...")
    wb.save(filename)
    
    logger.info("")
    logger.info("=" * 78)
    logger.info("\u2705 SUCCESS: Workbook generated successfully!")
    logger.info("=" * 78)
    logger.info("")
    logger.info(f"📄 Filename: {filename}")
    logger.info(f"📊 Estimated file size: ~700 KB - 1.2 MB")
    logger.info("")
    logger.info("Workbook Structure (14 sheets):")
    logger.info("  Y Sheet 1:  Instructions & Legend")
    logger.info("  Y Sheet 2:  SIEM Platform Details (22 component rows)")
    logger.info("  Y Sheet 3:  Log Forwarder Inventory (192 forwarder rows)")
    logger.info("  Y Sheet 4:  Collection Reliability (192 system rows)")
    logger.info("  Y Sheet 5:  Integration Architecture (42 integration points)")
    logger.info("  Y Sheet 6:  SIEM Storage & Capacity (12 storage tiers)")
    logger.info("  Y Sheet 7:  Log Parsing & Normalization (92 log sources)")
    logger.info("  Y Sheet 8:  SIEM Performance Metrics (82 daily rows)")
    logger.info("  Y Sheet 9:  Data Quality Assessment (42 quality checks)")
    logger.info("  Y Sheet 10: Encryption & Authentication (TLS, auth methods)")
    logger.info("  Y Sheet 11: Gap Analysis & Remediation (92 gap rows)")
    logger.info("  Y Sheet 12: Evidence Register (audit documentation)")
    logger.info("  Y Sheet 13: Summary Dashboard (with health metrics)")
    logger.info("  Y Sheet 14: Approval & Sign-Off")
    logger.info("")
    logger.info("Features:")
    logger.info("  ✓ Auto-generated IDs (Forwarder, Gap)")
    logger.info("  ✓ Collection reliability formulas")
    logger.info("  ✓ Storage capacity calculations")
    logger.info("  ✓ Performance status indicators")
    logger.info("  ✓ Data quality scoring")
    logger.info("  ✓ Conditional formatting (Green/Yellow/Red)")
    logger.info("  ✓ Date format: DD.MM.YYYY")
    logger.info("")
    logger.info("Next Steps:")
    logger.info("  1. Document SIEM platform architecture")
    logger.info("  2. Inventory all log forwarders")
    logger.info("  3. Collect 30 days of reliability metrics")
    logger.info("  4. Assess storage capacity and growth")
    logger.info("  5. Review parsing accuracy")
    logger.info("  6. Track performance metrics")
    logger.info("  7. Evaluate data quality")
    logger.info("  8. Document gaps and remediation plans")
    logger.info("")
    logger.info("═" * 78)
    logger.info("Remember: 'You can't manage what you don't measure.'")
    logger.info("Measure your log collection infrastructure properly!")
    logger.info("═" * 78)
    logger.info("")


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
