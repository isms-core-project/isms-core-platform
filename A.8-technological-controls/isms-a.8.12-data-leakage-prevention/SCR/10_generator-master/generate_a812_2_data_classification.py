#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.12.2 - Data Classification Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Annex A Control A.8.12: Data Leakage Prevention
Assessment Domain 2 of 4: Data Classification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific data classification schema, sensitive data
categories, and regulatory requirements.

Key customization areas:
1. Classification schema levels (match your data classification policy)
2. Sensitive data categories (PII, financial per your definitions)
3. Regulatory mappings (FADP, GDPR, PCI-DSS per your jurisdictions)
4. Data discovery tools (specific to your technology stack)
5. Ownership assignment criteria (aligned with your governance)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for DLP)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

Generates comprehensive data classification assessment workbook for systematic
evaluation of sensitive data identification and classification against ISO 27001:2022
Control A.8.12 requirements.

This workbook provides audit-ready evidence collection framework covering:
• Organizational data classification schema definition
• Sensitive data inventory (PII, financial, intellectual property, credentials)
• Data location mapping (file servers, databases, cloud storage, endpoints)
• Data ownership assignment and accountability
• Regulatory compliance mapping (Swiss FADP, EU GDPR, PCI-DSS)
• Data labeling methods and automated discovery tools
• Gap analysis and remediation planning
• Evidence register for audit traceability

--------------------------------------------------------------------------------
GENERATED WORKBOOK STRUCTURE
--------------------------------------------------------------------------------

Output File: ISMS-IMP-A.8.12.2_Data_Classification_YYYYMMDD.xlsx

Sheets (11 total):
1. Instructions_Legend - Assessment guidance and metadata
2. Classification_Schema - Organizational data classification levels
3. Sensitive_Data_Inventory - Complete inventory of sensitive data types
4. Data_Locations - Where sensitive data resides (systems, applications)
5. Data_Ownership - Data owners and custodians assignment
6. PII_Inventory - Personal Identifiable Information detailed assessment
7. Regulatory_Mapping - FADP, GDPR, PCI-DSS compliance requirements
8. Data_Discovery_Tools - Automated discovery and scanning tools
9. Gap_Analysis - Classification gaps requiring remediation (40 rows)
10. Evidence_Register - Audit evidence tracking (100 rows)
11. Summary_Dashboard - Compliance metrics and KPIs

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System:
    Python 3.8 or higher
    Ubuntu/Debian Linux (recommended) or macOS

Dependencies:
    openpyxl - Excel file generation library
    
Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic:
    python3 generate_a812_2_data_classification.py

Output Location:
    Current working directory
    
Output Filename:
    ISMS-IMP-A.8.12.2_Data_Classification_YYYYMMDD.xlsx
    (Where YYYYMMDD = current date)

Post-Generation:
    1. Open workbook in Microsoft Excel or LibreOffice Calc
    2. Complete all yellow input fields (organization-specific data)
    3. Review pre-populated examples (gray rows) for guidance
    4. Define your organization's data classification schema
    5. Inventory all sensitive data types and locations
    6. Assign data owners and map regulatory requirements
    7. Collect and document evidence (Evidence_Register sheet)
    8. Complete gap analysis for identified deficiencies
    9. Obtain management approval (Summary_Dashboard sheet)

--------------------------------------------------------------------------------
FRAMEWORK INTEGRATION
--------------------------------------------------------------------------------

Position in A.8.12 Framework:
    Assessment Domain 2 of 4 in comprehensive DLP evaluation framework
    Focus: Data classification, identification, and regulatory mapping
    
Related Documents:
    Policy:         ISMS-POL-A.8.12-S2.1 (Data Classification Requirements)
    Implementation: ISMS-IMP-A.8.12.2 (Data Classification Implementation Guide)
    Dashboard:      ISMS-IMP-A.8.12.5 (Compliance Dashboard)

Integration Workflow:
    1. Generate assessment workbooks:
       python3 generate_a812_1_dlp_infrastructure.py
       python3 generate_a812_2_data_classification.py      ← YOU ARE HERE
       python3 generate_a812_3_channel_coverage.py
       python3 generate_a812_4_monitoring_response.py
    
    2. Complete assessments (manual - security team, data owners, legal)
    
    3. Normalize filenames for dashboard linking:
       python3 normalize_assessment_files_a812.py
    
    4. Generate executive dashboard:
       python3 generate_a812_5_compliance_dashboard.py
    
    5. Consolidate assessment data:
       python3 consolidate_a812_dashboard.py [dashboard_file]
    
    6. Present to CISO/auditors (evidence-based compliance reporting)

Data Flow:
    THIS SCRIPT → Classification Assessment → Normalize → Dashboard → Audit Evidence

Critical Prerequisite:
    Complete this assessment BEFORE completing Domain 3 (Channel Coverage).
    Channel DLP policies depend on data classification taxonomy defined here.

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.12
Assessment Domain:    2 of 4 (Data Classification)
Framework Version:    1.0
Script Version:       1.0
Date:                 [Date to be set]
Author:               [Organization] ISMS Implementation Team
License:              [Organization License/Terms]

Related Standards:
    - ISO/IEC 27002:2022 (Information Security Controls)
    - Swiss FADP (Federal Act on Data Protection - Bundesgesetz über den Datenschutz)
    - EU GDPR (General Data Protection Regulation)
    - PCI-DSS v4.0 (Payment Card Industry Data Security Standard)
    - NIST SP 800-53 (Security and Privacy Controls)
    - CIS Controls v8 (Center for Internet Security)

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

Data Classification Prerequisites:
    Before completing this assessment, ensure you have:
    • Executive approval for data classification schema
    • Identified key data owners across organization
    • Understanding of applicable regulatory requirements (FADP, GDPR, etc.)
    • Inventory of data storage systems (file servers, databases, cloud)
    • Access to data discovery tool results (if available)

Swiss FADP Compliance:
    This assessment includes Swiss Federal Act on Data Protection (FADP) requirements.
    Key focus areas:
    • Personal data identification and categorization
    • Special category personal data (sensitive data)
    • Data subject rights (access, deletion, portability)
    • Cross-border data transfer requirements
    • Data breach notification obligations

Assessment Scope:
    This workbook covers ALL sensitive data categories:
    • Personal Identifiable Information (PII) - names, addresses, contact details
    • Financial data - credit cards, bank accounts, payment information
    • Intellectual property - trade secrets, proprietary algorithms, R&D data
    • Credentials - passwords, API keys, certificates, tokens
    • Health data - medical records, health insurance information
    • Special category data - biometrics, genetic data, political opinions

Data Classification:
    Generated workbooks contain sensitive organizational security information.
    Handle according to [Organization]'s data classification policy.
    Recommended classification: [Organization] Confidential

Audit Considerations:
    This workbook generates ISO 27001:2022 audit evidence per Control A.8.12.
    Ensure all fields completed accurately and evidence properly documented.
    Retain completed workbooks for audit cycle (typically 3 years).
    Auditors will verify: classification schema, data inventory, regulatory mapping.

Review Cycle:
    Quarterly: Update sensitive data inventory and locations
    Annually: Complete full data classification reassessment
    Ad-hoc: When new data types identified or regulations change

================================================================================
END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.12.2"
WORKBOOK_NAME = "Data Classification Assessment"
CONTROL_ID = "A.8.12"
CONTROL_NAME = "Data Leakage Prevention"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

CHECK_MARK = "\u2705"      # ✅
CROSS_MARK = "\u274C"      # ❌
WARNING = "\u26A0"         # ⚠️
CLIPBOARD = "\u1F4CB"      # 📋
TRIANGLE = "\u25B8"        # ▸
BULLET = "\u2022"          # •



# ============================================================================
# SECTION 1: CONSTANTS & CONFIGURATION
# ============================================================================

# Document Information
WORKBOOK_VERSION = "1.0"
CONTROL_ID = "A.8.12"
WORKBOOK_ID = "ISMS-IMP-A.8.12.2"
RELATED_POLICY = "ISMS-POL-A.8.12-S2.1"
ASSESSMENT_AREA = "Data Classification"

# Color Scheme (CONSISTENT across all A.8.12 workbooks)
COLOR_HEADER = "003366"          # Dark blue
COLOR_SUBHEADER = "4472C4"       # Medium blue
COLOR_COLUMN_HEADER = "D9D9D9"   # Light gray
COLOR_INPUT = "FFFFCC"           # Light yellow (user input)
COLOR_INFO = "E7E6E6"            # Light gray (info/example rows)
COLOR_COMPLETE = "C6EFCE"        # Light green (\u2705 Yes)
COLOR_PARTIAL = "FFEB9C"         # Light yellow (\u26A0\uFE0F Partial)
COLOR_MISSING = "FFC7CE"         # Light red (\u274C No)
COLOR_PLANNED = "B4C7E7"         # Light blue (\u1F4CB Planned)

# Standard column widths
WIDTH_NARROW = 12
WIDTH_MEDIUM = 20
WIDTH_WIDE = 25
WIDTH_EXTRA_WIDE = 30
WIDTH_DESCRIPTION = 35
WIDTH_VERY_WIDE = 40


# ============================================================================
# SECTION 2: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create all sheets in order
    sheets = [
        "Instructions_Legend",
        "Classification_Schema",
        "Sensitive_Data_Inventory",
        "Data_Location_Mapping",
        "Data_Owner_Assignment",
        "Regulatory_Mapping",
        "Labeling_Methods",
        "Discovery_Results",
        "Gap_Analysis",
        "Evidence_Register",
        "Summary_Dashboard",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    
    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style TEMPLATES (dictionaries), not reusable objects.
    
    CRITICAL: Do NOT create shared Font/Fill/Border objects.
    Each cell gets its OWN style instance to avoid openpyxl issues.
    """
    return {
        "header": {
            "font": Font(name="Arial", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "subheader": {
            "font": Font(name="Arial", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "column_header": {
            "font": Font(name="Arial", size=10, bold=True),
            "fill": PatternFill(start_color=COLOR_COLUMN_HEADER, end_color=COLOR_COLUMN_HEADER, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "input_cell": {
            "fill": PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "info_cell": {
            "fill": PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        },
    }


def apply_style(cell, style_dict):
    """
    Apply style template to cell by creating NEW instances.
    NEVER reuse Font/Fill/Border objects across cells.
    """
    if "font" in style_dict:
        cell.font = Font(**{k: v for k, v in style_dict["font"].__dict__.items() if not k.startswith('_')})
    if "fill" in style_dict:
        cell.fill = PatternFill(**{k: v for k, v in style_dict["fill"].__dict__.items() if not k.startswith('_')})
    if "alignment" in style_dict:
        cell.alignment = Alignment(**{k: v for k, v in style_dict["alignment"].__dict__.items() if not k.startswith('_')})
    if "border" in style_dict:
        cell.border = Border(**{k: v for k, v in style_dict["border"].__dict__.items() if not k.startswith('_')})


# ============================================================================
# SECTION 3: DATA VALIDATIONS
# ============================================================================

def create_data_validations():
    """
    Create data validation objects.
    MUST be added to worksheet.add_data_validation() and then cells added to validation.
    """
    return {
        "yes_no_partial": DataValidation(
            type="list",
            formula1='"Yes,No,Partial,Planned,N/A"',
            allow_blank=False,
            showDropDown=True,
            showErrorMessage=True,
            error="Invalid value. Select from dropdown.",
            errorTitle="Invalid Entry"
        ),
        "yes_no": DataValidation(
            type="list",
            formula1='"Yes,No,N/A"',
            allow_blank=False,
            showDropDown=True
        ),
        "yes_no_pending": DataValidation(
            type="list",
            formula1='"Yes,No,Pending"',
            allow_blank=False,
            showDropDown=True
        ),
        "classification_level": DataValidation(
            type="list",
            formula1='"Public,Internal,Confidential,Restricted"',
            allow_blank=False,
            showDropDown=True
        ),
        "data_type": DataValidation(
            type="list",
            formula1='"PII,Financial,IP,Credentials,Business Confidential,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "regulatory": DataValidation(
            type="list",
            formula1='"FADP,GDPR,PCI-DSS,FADP/GDPR,Multiple,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "discovery_status": DataValidation(
            type="list",
            formula1='"Discovered,In Progress,Not Started"',
            allow_blank=False,
            showDropDown=True
        ),
        "location_type": DataValidation(
            type="list",
            formula1='"File Server,Database,Endpoint,Cloud,Email,Backup,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "labeling_method": DataValidation(
            type="list",
            formula1='"Manual,Automated,Metadata,None"',
            allow_blank=False,
            showDropDown=True
        ),
        "compliance_status": DataValidation(
            type="list",
            formula1='"Compliant,Non-Compliant,Partial,Planned,N/A"',
            allow_blank=False,
            showDropDown=True
        ),
        "remediation_status": DataValidation(
            type="list",
            formula1='"Complete,In Progress,Planned,Not Started"',
            allow_blank=False,
            showDropDown=True
        ),
        "risk_level": DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False,
            showDropDown=True
        ),
        "gap_status": DataValidation(
            type="list",
            formula1='"Open,In Progress,Resolved,Accepted,Closed"',
            allow_blank=False,
            showDropDown=True
        ),
        "evidence_type": DataValidation(
            type="list",
            formula1='"Screenshot,Configuration File,Policy Document,Log Export,Report,Certificate,Email,Meeting Minutes,Other"',
            allow_blank=False,
            showDropDown=True
        ),
        "verification_status": DataValidation(
            type="list",
            formula1='"Verified,Pending,Rejected"',
            allow_blank=False,
            showDropDown=True
        ),
        "approval_status": DataValidation(
            type="list",
            formula1='"Approved,Pending,Rejected"',
            allow_blank=False,
            showDropDown=True
        ),
    }

# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS - PART 1
# ============================================================================

def create_instructions(ws, styles):
    """Create Instructions & Legend sheet."""
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = f"ISMS Control {WORKBOOK_ID} - {ASSESSMENT_AREA}"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 40
    
    # Subheader
    ws.merge_cells('A2:H2')
    ws['A2'] = f"ISO/IEC 27001:2022 Control {CONTROL_ID} - Data Leakage Prevention"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Document Info
    info = [
        ("Document ID:", WORKBOOK_ID),
        ("Assessment Area:", ASSESSMENT_AREA),
        ("Related Policy:", RELATED_POLICY),
        ("Version:", WORKBOOK_VERSION),
        ("Assessment Date:", "[USER INPUT]"),
        ("Completed By:", "[USER INPUT]"),
        ("Organization:", "[USER INPUT]"),
        ("Review Cycle:", "Quarterly"),
    ]
    
    row = 4
    for label, value in info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if "USER INPUT" in value:
            ws[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
        row += 1
    
    # HOW TO USE THIS WORKBOOK section
    row += 2
    ws[f'A{row}'] = "HOW TO USE THIS WORKBOOK"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    instructions = [
        "1. Complete each worksheet tab in sequence (Schema → Inventory → Locations → Owners → ...)",
        "2. Fill ALL yellow-highlighted cells with your organization's information",
        "3. Use dropdown menus where provided (do not type free-form text in dropdown cells)",
        "4. Document your data classification schema (minimum 4 levels: Public/Internal/Confidential/Restricted)",
        "5. Inventory ALL sensitive data categories (PII, financial, IP, credentials, business confidential)",
        "6. Map data to storage locations (file servers, databases, cloud, endpoints)",
        "7. Assign data owners and stewards for accountability",
        "8. Map data categories to regulatory requirements (FADP, GDPR, PCI-DSS)",
        "9. Provide evidence IDs for all assessments (format: A812-2-[CATEGORY]-[###])",
        "10. Review Summary Dashboard for overall classification compliance score",
        "11. Obtain DPO/CISO approval before finalizing",
    ]
    
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # LEGEND section
    row += 2
    ws[f'A{row}'] = "LEGEND - RESPONSE VALUES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    legend = [
        ("Yes", "Fully implemented and documented", COLOR_COMPLETE),
        ("No", "Not implemented", COLOR_MISSING),
        ("Partial", "Partially implemented (explain in notes)", COLOR_PARTIAL),
        ("Planned", "Scheduled for implementation (provide target date)", COLOR_PLANNED),
        ("N/A", "Not applicable (provide justification)", "FFFFFF"),
    ]
    
    for value, description, color in legend:
        ws[f'A{row}'] = value
        ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = description
        row += 1
    
    # CLASSIFICATION LEVELS section
    row += 2
    ws[f'A{row}'] = "DATA CLASSIFICATION LEVELS (Standard)"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    classification_info = [
        ("Public", "Information approved for public disclosure"),
        ("Internal", "Information for internal use only, no external sharing"),
        ("Confidential", "Sensitive information requiring protection, limited internal access"),
        ("Restricted", "Highly sensitive data, need-to-know basis, strict controls"),
    ]
    
    for level, description in classification_info:
        ws[f'A{row}'] = level
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = description
        row += 1
    
    # EVIDENCE NAMING section
    row += 2
    ws[f'A{row}'] = "EVIDENCE NAMING CONVENTION"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    evidence_examples = [
        "Format: A812-2-[CATEGORY]-[###]",
        "Examples:",
        "  \u2022 A812-2-CLS-001 = Classification - Schema document",
        "  \u2022 A812-2-INV-001 = Inventory - Data discovery report",
        "  \u2022 A812-2-OWN-001 = Ownership - Assignment matrix",
        "  \u2022 A812-2-REG-001 = Regulatory - FADP compliance mapping",
        "  \u2022 A812-2-LBL-001 = Labeling - AIP configuration screenshot",
    ]
    
    for example in evidence_examples:
        ws[f'A{row}'] = example
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_MEDIUM
    ws.column_dimensions['B'].width = WIDTH_EXTRA_WIDE
    
    # Freeze panes at row 3
    ws.freeze_panes = 'A3'


def create_classification_schema(ws, styles):
    """Create Classification Schema sheet."""
    
    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "DATA CLASSIFICATION SCHEMA"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:J2')
    ws['A2'] = "Document your organizational data classification levels and handling requirements"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Classification Level"),
        ("B3", "Definition"),
        ("C3", "Examples"),
        ("D3", "Handling Requirements"),
        ("E3", "Access Controls"),
        ("F3", "Encryption Required"),
        ("G3", "DLP Monitoring"),
        ("H3", "Retention Period"),
        ("I3", "Disposal Method"),
        ("J3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [20, WIDTH_VERY_WIDE, WIDTH_VERY_WIDE, WIDTH_DESCRIPTION, WIDTH_EXTRA_WIDE, 18, 18, 18, 25, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated classification levels (info rows - partial user input)
    levels = [
        ("Public", "[Define what Public means for your organization]", "[Examples: Press releases, marketing materials, public website]", 
         "[Handling: Can be shared externally]", "[Access: No restrictions]"),
        ("Internal", "[Define Internal classification]", "[Examples: Internal policies, procedures, org charts]",
         "[Handling: Internal use only, no external sharing without approval]", "[Access: All employees]"),
        ("Confidential", "[Define Confidential classification]", "[Examples: Customer lists, financial reports, contracts]",
         "[Handling: Limited distribution, need-to-know basis]", "[Access: Authorized personnel only]"),
        ("Restricted", "[Define Restricted classification]", "[Examples: Trade secrets, M&A, executive comp, PII]",
         "[Handling: Strictly controlled, encrypted, logged access]", "[Access: Explicit authorization required]"),
    ]
    
    row = 4
    for level_data in levels:
        ws[f'A{row}'] = level_data[0]
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
        
        # Columns B-E are user input (yellow)
        for col_idx, value in enumerate(level_data[1:], start=2):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["input_cell"])
        
        # Columns F-J are dropdowns/input (yellow)
        for col_idx in range(6, 11):
            cell = ws.cell(row=row, column=col_idx)
            apply_style(cell, styles["input_cell"])
        
        row += 1
    
    # Additional blank rows for custom classification levels
    for r in range(row, row + 6):
        for col_idx in range(1, 11):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Add data validations
    validations = create_data_validations()
    
    # Encryption Required (Column F)
    ws.add_data_validation(validations['yes_no'])
    for r in range(4, 14):
        validations['yes_no'].add(ws[f'F{r}'])
    
    # DLP Monitoring (Column G)
    ws.add_data_validation(validations['yes_no_partial'])
    for r in range(4, 14):
        validations['yes_no_partial'].add(ws[f'G{r}'])
    
    ws.freeze_panes = 'A4'


def create_sensitive_data_inventory(ws, styles):
    """Create Sensitive Data Inventory sheet."""
    
    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "SENSITIVE DATA INVENTORY"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:K2')
    ws['A2'] = "Inventory ALL sensitive data categories requiring DLP protection"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Data Category"),
        ("B3", "Data Type"),
        ("C3", "Classification Level"),
        ("D3", "Regulatory Requirement"),
        ("E3", "Data Examples"),
        ("F3", "Business Owner"),
        ("G3", "Data Steward"),
        ("H3", "Estimated Volume"),
        ("I3", "Discovery Status"),
        ("J3", "DLP Protection"),
        ("K3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [30, 25, 20, 20, WIDTH_DESCRIPTION, 25, 25, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated examples (gray rows)
    examples = [
        ("Personal Names", "PII", "Confidential", "FADP/GDPR", "First name, last name, full name", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Yes", "A812-2-INV-001"),
        ("Email Addresses", "PII", "Confidential", "FADP/GDPR", "john.doe@example.com", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Yes", "A812-2-INV-002"),
        ("Swiss Social Security (AHV)", "PII", "Restricted", "FADP", "756.1234.5678.90", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Yes", "A812-2-INV-003"),
        ("Credit Card Numbers", "Financial", "Restricted", "PCI-DSS", "4111-1111-1111-1111", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Yes", "A812-2-INV-004"),
        ("Bank Account (IBAN)", "Financial", "Restricted", "FADP/GDPR", "CH93 0076 2011 6238 5295 7", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Partial", "A812-2-INV-005"),
        ("API Keys", "Credentials", "Restricted", "None", "sk_live_abcd1234efgh5678", "[Owner]", "[Steward]", "[Volume]", "In Progress", "Partial", "A812-2-INV-006"),
        ("Source Code", "IP", "Confidential", "None", "Proprietary algorithms, trade secrets", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Yes", "A812-2-INV-007"),
        ("Customer Lists", "Business Confidential", "Confidential", "FADP/GDPR", "CRM exports, contact databases", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Partial", "A812-2-INV-008"),
        ("M&A Documents", "Business Confidential", "Restricted", "None", "Acquisition targets, financial models", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Yes", "A812-2-INV-009"),
        ("Employee Health Records", "PII", "Restricted", "FADP/GDPR", "Medical diagnoses, disability status", "[Owner]", "[Steward]", "[Volume]", "Discovered", "Partial", "A812-2-INV-010"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 20):
        for col_idx in range(1, 12):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Add data validations
    validations = create_data_validations()
    
    # Data Type (Column B)
    ws.add_data_validation(validations['data_type'])
    for r in range(14, 34):
        validations['data_type'].add(ws[f'B{r}'])
    
    # Classification Level (Column C)
    ws.add_data_validation(validations['classification_level'])
    for r in range(14, 34):
        validations['classification_level'].add(ws[f'C{r}'])
    
    # Regulatory Requirement (Column D)
    ws.add_data_validation(validations['regulatory'])
    for r in range(14, 34):
        validations['regulatory'].add(ws[f'D{r}'])
    
    # Discovery Status (Column I)
    ws.add_data_validation(validations['discovery_status'])
    for r in range(14, 34):
        validations['discovery_status'].add(ws[f'I{r}'])
    
    # DLP Protection (Column J)
    ws.add_data_validation(validations['yes_no_partial'])
    for r in range(14, 34):
        validations['yes_no_partial'].add(ws[f'J{r}'])
    
    ws.freeze_panes = 'A4'


def create_data_location_mapping(ws, styles):
    """Create Data Location Mapping sheet."""
    
    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "DATA LOCATION MAPPING"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:K2')
    ws['A2'] = "Map sensitive data categories to storage locations (file servers, databases, cloud, endpoints)"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Data Category"),
        ("B3", "Storage Location Type"),
        ("C3", "Specific Location"),
        ("D3", "Path/Schema"),
        ("E3", "Estimated Records/Files"),
        ("F3", "Last Discovery Scan"),
        ("G3", "DLP Coverage"),
        ("H3", "Encryption at Rest"),
        ("I3", "Access Controls"),
        ("J3", "Data Owner Notified"),
        ("K3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [30, 20, 30, WIDTH_DESCRIPTION, 18, 15, 18, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated examples
    examples = [
        ("Credit Card Numbers", "Database", "SQL-PROD-01", "dbo.Payments.CCNumber", "250000", "2024-12-10", "Yes", "Yes", "Yes", "Yes", "A812-2-LOC-001"),
        ("Customer PII", "File Server", "FS-HR-01", "\\\\HR\\PersonnelFiles\\", "5000", "2024-12-15", "Partial", "Yes", "Yes", "Yes", "A812-2-LOC-002"),
        ("Source Code", "Cloud", "GitHub Enterprise", "/repos/proprietary/", "150", "2024-12-01", "Yes", "Yes", "Yes", "Yes", "A812-2-LOC-003"),
        ("API Keys", "Endpoint", "Developer Workstations", "Various local configs", "[Unknown]", "2024-11-20", "No", "Partial", "Partial", "No", "A812-2-LOC-004"),
        ("Email Archives", "Email", "M365 Exchange Online", "All mailboxes", "2000000", "2024-12-20", "Yes", "Yes", "Yes", "Yes", "A812-2-LOC-005"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 35):
        for col_idx in range(1, 12):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Add data validations
    validations = create_data_validations()
    
    # Storage Location Type (Column B)
    ws.add_data_validation(validations['location_type'])
    for r in range(9, 45):
        validations['location_type'].add(ws[f'B{r}'])
    
    # DLP Coverage, Encryption, Access Controls (Columns G, H, I)
    for col in ['G', 'H', 'I']:
        ws.add_data_validation(validations['yes_no_partial'])
        for r in range(9, 45):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])
    
    # Data Owner Notified (Column J)
    ws.add_data_validation(validations['yes_no_pending'])
    for r in range(9, 45):
        validations['yes_no_pending'].add(ws[f'J{r}'])
    
    ws.freeze_panes = 'A4'

# ============================================================================
# SECTION 5: SHEET CREATION FUNCTIONS - PART 2
# ============================================================================

def create_data_owner_assignment(ws, styles):
    """Create Data Owner Assignment sheet."""
    
    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "DATA OWNER ASSIGNMENT"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:K2')
    ws['A2'] = "Assign data owners and stewards for each data category - accountability is key!"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Data Category"),
        ("B3", "Business Owner Name"),
        ("C3", "Business Owner Department"),
        ("D3", "Business Owner Email"),
        ("E3", "Data Steward Name"),
        ("F3", "Data Steward Department"),
        ("G3", "Data Steward Email"),
        ("H3", "Ownership Documented"),
        ("I3", "Owner Training Complete"),
        ("J3", "Last Review Date"),
        ("K3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [30, 25, 20, 30, 25, 20, 30, 18, 18, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated examples
    examples = [
        ("Customer PII", "Jane Smith", "Sales", "jane.smith@example.com", "John Security", "IT Security", "john.sec@example.com", "Yes", "Yes", "2024-12-01", "A812-2-OWN-001"),
        ("Financial Data", "Robert Johnson", "Finance", "robert.j@example.com", "Mary InfoSec", "IT Security", "mary.is@example.com", "Yes", "Partial", "2024-11-15", "A812-2-OWN-002"),
        ("Source Code", "Alice Developer", "R&D", "alice.dev@example.com", "Bob CISO", "IT Security", "bob.ciso@example.com", "Yes", "Yes", "2024-12-10", "A812-2-OWN-003"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 27):
        for col_idx in range(1, 12):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Add data validations
    validations = create_data_validations()
    
    # Ownership Documented (Column H)
    ws.add_data_validation(validations['yes_no_pending'])
    for r in range(7, 34):
        validations['yes_no_pending'].add(ws[f'H{r}'])
    
    # Owner Training Complete (Column I)
    ws.add_data_validation(validations['yes_no_partial'])
    for r in range(7, 34):
        validations['yes_no_partial'].add(ws[f'I{r}'])
    
    ws.freeze_panes = 'A4'


def create_regulatory_mapping(ws, styles):
    """Create Regulatory Mapping sheet."""
    
    # Header
    ws.merge_cells('A1:I1')
    ws['A1'] = "REGULATORY COMPLIANCE MAPPING"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:I2')
    ws['A2'] = "Map data categories to regulatory requirements (FADP, GDPR, PCI-DSS, etc.)"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Data Category"),
        ("B3", "Regulation"),
        ("C3", "Specific Article/Section"),
        ("D3", "Requirement Summary"),
        ("E3", "Compliance Status"),
        ("F3", "DLP Controls Required"),
        ("G3", "Breach Notification Required"),
        ("H3", "Data Subject Rights"),
        ("I3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [30, 20, 25, WIDTH_VERY_WIDE, 18, 18, 18, 30, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated examples
    examples = [
        ("Personal Names", "FADP", "Article 4(1)", "Defined as personal data, requires appropriate protection measures", "Compliant", "Yes", "Yes", "Access, deletion, portability", "A812-2-REG-001"),
        ("Special Category PII", "GDPR", "Article 9", "Sensitive data (health, biometric, etc.) - strict processing rules", "Compliant", "Yes", "Yes", "All GDPR rights apply", "A812-2-REG-002"),
        ("Credit Card Numbers", "PCI-DSS", "Requirement 3", "Protect stored cardholder data with encryption and access controls", "Partial", "Yes", "Yes", "N/A (PCI-DSS)", "A812-2-REG-003"),
        ("Swiss AHV Number", "FADP", "Article 5", "Special category personal data, heightened protection required", "Compliant", "Yes", "Yes", "Access, deletion, correction", "A812-2-REG-004"),
        ("Employee Health Data", "FADP/GDPR", "FADP Art 5, GDPR Art 9", "Health data is special category, strict consent and security", "Compliant", "Yes", "Yes", "All data subject rights", "A812-2-REG-005"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 20):
        for col_idx in range(1, 10):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Add data validations
    validations = create_data_validations()
    
    # Regulation (Column B)
    regulation_val = DataValidation(
        type="list",
        formula1='"FADP,GDPR,PCI-DSS,HIPAA,SOX,FADP/GDPR,Multiple,Other,None"',
        allow_blank=False
    )
    ws.add_data_validation(regulation_val)
    for r in range(9, 29):
        regulation_val.add(ws[f'B{r}'])
    
    # Compliance Status (Column E)
    ws.add_data_validation(validations['compliance_status'])
    for r in range(9, 29):
        validations['compliance_status'].add(ws[f'E{r}'])
    
    # DLP Controls Required (Column F)
    ws.add_data_validation(validations['yes_no_partial'])
    for r in range(9, 29):
        validations['yes_no_partial'].add(ws[f'F{r}'])
    
    # Breach Notification Required (Column G)
    ws.add_data_validation(validations['yes_no'])
    for r in range(9, 29):
        validations['yes_no'].add(ws[f'G{r}'])
    
    ws.freeze_panes = 'A4'


def create_labeling_methods(ws, styles):
    """Create Labeling Methods sheet."""
    
    # Header
    ws.merge_cells('A1:I1')
    ws['A1'] = "DATA LABELING METHODS"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:I2')
    ws['A2'] = "Assess data labeling and classification marking methods across systems"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "System/Application"),
        ("B3", "Labeling Method"),
        ("C3", "Classification Tool"),
        ("D3", "Supported Labels"),
        ("E3", "User Training Provided"),
        ("F3", "Enforcement Capability"),
        ("G3", "DLP Integration"),
        ("H3", "Adoption Rate %"),
        ("I3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [25, 20, 25, 30, 18, 18, 18, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated examples
    examples = [
        ("Microsoft 365", "Automated", "Microsoft Purview (AIP)", "Public, Internal, Confidential, Restricted", "Yes", "Yes", "Yes", "85", "A812-2-LBL-001"),
        ("SharePoint", "Manual", "Built-in Classification", "Public, Internal, Restricted", "Partial", "Partial", "Partial", "60", "A812-2-LBL-002"),
        ("File Servers", "None", "None", "None", "No", "No", "No", "0", "A812-2-LBL-003"),
        ("Email (Outlook)", "Manual", "Azure Information Protection", "All 4 levels", "Yes", "Yes", "Yes", "70", "A812-2-LBL-004"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 16):
        for col_idx in range(1, 10):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Add data validations
    validations = create_data_validations()
    
    # Labeling Method (Column B)
    ws.add_data_validation(validations['labeling_method'])
    for r in range(8, 24):
        validations['labeling_method'].add(ws[f'B{r}'])
    
    # User Training, Enforcement, DLP Integration (Columns E, F, G)
    for col in ['E', 'F', 'G']:
        ws.add_data_validation(validations['yes_no_partial'])
        for r in range(8, 24):
            validations['yes_no_partial'].add(ws[f'{col}{r}'])
    
    ws.freeze_panes = 'A4'


def create_discovery_results(ws, styles):
    """Create Discovery Results sheet."""
    
    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "DATA DISCOVERY RESULTS"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:J2')
    ws['A2'] = "Document data discovery and scanning results (Spirion, BigID, manual audits)"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Discovery Tool"),
        ("B3", "Scan Target"),
        ("C3", "Scan Date"),
        ("D3", "Data Categories Found"),
        ("E3", "Total Findings"),
        ("F3", "Critical Findings"),
        ("G3", "False Positive Rate %"),
        ("H3", "Remediation Status"),
        ("I3", "Data Owner Notified"),
        ("J3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [25, 30, 15, 30, 15, 15, 15, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated examples
    examples = [
        ("Spirion", "FS-HR-01", "2024-12-15", "PII, SSN, AHV", "5243", "1200", "15", "In Progress", "Yes", "A812-2-DSC-001"),
        ("BigID", "SQL-PROD-01", "2024-12-10", "PII, CCN, IBAN", "250000", "50000", "5", "In Progress", "Yes", "A812-2-DSC-002"),
        ("Manual Audit", "SharePoint", "2024-11-20", "Business Confidential", "1500", "300", "30", "Complete", "Yes", "A812-2-DSC-003"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 27):
        for col_idx in range(1, 11):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Add data validations
    validations = create_data_validations()
    
    # Remediation Status (Column H)
    ws.add_data_validation(validations['remediation_status'])
    for r in range(7, 34):
        validations['remediation_status'].add(ws[f'H{r}'])
    
    # Data Owner Notified (Column I)
    ws.add_data_validation(validations['yes_no_pending'])
    for r in range(7, 34):
        validations['yes_no_pending'].add(ws[f'I{r}'])
    
    ws.freeze_panes = 'A4'


def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet (standard across all workbooks)."""
    
    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "GAP ANALYSIS - IDENTIFIED DEFICIENCIES"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:K2')
    ws['A2'] = "Document all data classification gaps and remediation plans"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Gap ID"),
        ("B3", "Gap Description"),
        ("C3", "Affected Data Category"),
        ("D3", "Risk Level"),
        ("E3", "Business Impact"),
        ("F3", "Root Cause"),
        ("G3", "Remediation Plan"),
        ("H3", "Owner"),
        ("I3", "Target Date"),
        ("J3", "Status"),
        ("K3", "Evidence ID"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [18, WIDTH_DESCRIPTION, 25, 15, 30, 30, WIDTH_DESCRIPTION, 20, 15, 15, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Pre-populated example gaps
    examples = [
        ("GAP-A812-2-001", "No data classification schema documented", "All data", "Critical", "Uncontrolled data handling, no DLP baseline", "Schema never formally defined", "Create and approve 4-tier classification schema", "[Owner]", "[Date]", "Open", "A812-2-GAP-001"),
        ("GAP-A812-2-002", "PII found on file servers without DLP protection", "Customer PII", "High", "Potential FADP/GDPR violation, unmonitored exfiltration", "File servers not in DLP scope", "Deploy endpoint DLP to file servers", "[Owner]", "[Date]", "Open", "A812-2-GAP-002"),
        ("GAP-A812-2-003", "Data owners not assigned for 40% of data categories", "Multiple", "High", "Lack of accountability, unclear data stewardship", "Ownership process not established", "Execute data owner assignment workshops", "[Owner]", "[Date]", "Open", "A812-2-GAP-003"),
    ]
    
    row = 4
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, styles["info_cell"])
        row += 1
    
    # User input rows
    for r in range(row, row + 37):
        for col_idx in range(1, 12):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Data validations
    validations = create_data_validations()
    
    ws.add_data_validation(validations['risk_level'])
    for r in range(7, 44):
        validations['risk_level'].add(ws[f'D{r}'])
    
    ws.add_data_validation(validations['gap_status'])
    for r in range(7, 44):
        validations['gap_status'].add(ws[f'J{r}'])
    
    ws.freeze_panes = 'A4'

# ============================================================================
# SECTION 6: SHEET CREATION FUNCTIONS - PART 3 (Standard Sheets)
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet (standard across all workbooks)."""
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = "EVIDENCE REGISTER - AUDIT TRAIL"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 30
    
    # Subheader
    ws.merge_cells('A2:H2')
    ws['A2'] = "Track all evidence collected during data classification assessment"
    apply_style(ws['A2'], styles["subheader"])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        ("A3", "Evidence ID"),
        ("B3", "Evidence Type"),
        ("C3", "Description"),
        ("D3", "Location/Link"),
        ("E3", "Date Collected"),
        ("F3", "Collected By"),
        ("G3", "Related Requirement"),
        ("H3", "Verification Status"),
    ]
    
    for cell, header in headers:
        ws[cell] = header
        apply_style(ws[cell], styles["column_header"])
    
    # Column widths
    widths = [18, 20, WIDTH_DESCRIPTION, 30, 15, 20, 25, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # User input rows (100 rows)
    for r in range(4, 104):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, styles["input_cell"])
    
    # Data validations
    validations = create_data_validations()
    
    ws.add_data_validation(validations['evidence_type'])
    for r in range(4, 104):
        validations['evidence_type'].add(ws[f'B{r}'])
    
    ws.add_data_validation(validations['verification_status'])
    for r in range(4, 104):
        validations['verification_status'].add(ws[f'H{r}'])
    
    ws.freeze_panes = 'A4'


def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard sheet with KPIs and compliance scoring."""
    
    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{ASSESSMENT_AREA} - COMPLIANCE DASHBOARD"
    apply_style(ws['A1'], styles["header"])
    ws.row_dimensions[1].height = 35
    
    # Document info section
    ws['A3'] = "Assessment Date:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = "[USER INPUT]"
    ws['B3'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    
    ws['A4'] = "Completed By:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "[USER INPUT]"
    ws['B4'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    
    ws['D3'] = "DPO Approval:"
    ws['D3'].font = Font(bold=True)
    ws['E3'] = "[Dropdown]"
    ws['E3'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    
    ws['D4'] = "Last Review:"
    ws['D4'].font = Font(bold=True)
    ws['E4'] = "[Date]"
    ws['E4'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    
    # KPI Section
    ws.merge_cells('A6:F6')
    ws['A6'] = "KEY PERFORMANCE INDICATORS"
    apply_style(ws['A6'], styles["subheader"])
    ws.row_dimensions[6].height = 25
    
    # KPI Headers
    kpi_headers = ['KPI', 'Current Value', 'Target', 'Status', 'Trend', 'Notes']
    for col_idx, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        apply_style(cell, styles["column_header"])
    
    # KPIs with formulas
    kpis = [
        ("Classification Schema Documented", '=COUNTIF(Classification_Schema!F$4:F$10,"Yes")', "4", '=IF(B8>=4,"\u2705 Pass","\u274C Fail")'),
        ("Sensitive Data Categories Inventoried", '=COUNTA(Sensitive_Data_Inventory!A$4:A$33)-10', "≥20", '=IF(B9>=20,"\u2705 Pass","\u274C Fail")'),
        ("Data Locations Mapped", '=COUNTA(Data_Location_Mapping!A$6:A$45)-5', "≥30", '=IF(B10>=30,"\u2705 Pass","\u274C Fail")'),
        ("Data Owners Assigned %", '=COUNTIF(Data_Owner_Assignment!H$4:H$33,"Yes")/B9*100', "≥90%", '=IF(B11>=90,"\u2705 Pass","\u274C Fail")'),
        ("DLP Protection Coverage %", '=COUNTIF(Sensitive_Data_Inventory!J$4:J$33,"Yes")/B9*100', "≥80%", '=IF(B12>=80,"\u2705 Pass","\u274C Fail")'),
        ("Regulatory Compliance %", '=COUNTIF(Regulatory_Mapping!E$4:E$28,"Compliant")/(COUNTA(Regulatory_Mapping!A$4:A$28)-5)*100', "≥90%", '=IF(B13>=90,"\u2705 Pass","\u274C Fail")'),
        ("Labeling Tool Adoption %", '=AVERAGE(Labeling_Methods!H$4:H$23)', "≥75%", '=IF(B14>=75,"\u2705 Pass","\u274C Fail")'),
        ("Data Discovery Coverage %", '=COUNTIF(Discovery_Results!H$4:H$33,"Complete")/(COUNTA(Discovery_Results!A$4:A$33)-3)*100', "≥80%", '=IF(B15>=80,"\u2705 Pass","\u274C Fail")'),
        ("Total Gaps Identified", '=COUNTA(Gap_Analysis!A$4:A$43)-3', "≤10", '=IF(B16<=10,"\u2705 Pass","\u274C Fail")'),
        ("Critical/High Gaps", '=COUNTIFS(Gap_Analysis!D$4:D$43,"Critical")+COUNTIFS(Gap_Analysis!D$4:D$43,"High")', "0", '=IF(B17=0,"\u2705 Pass","\u274C Fail")'),
        ("Overall Classification Compliance %", '=AVERAGE(B11:B15)', "≥85%", '=IF(B18>=85,"\u2705 Compliant",IF(B18>=70,"\u26A0\uFE0F Partial","\u274C Non-Compliant"))'),
    ]
    
    row = 8
    for kpi_name, formula, target, status_formula in kpis:
        ws[f'A{row}'] = kpi_name
        ws[f'B{row}'] = formula
        ws[f'C{row}'] = target
        ws[f'D{row}'] = status_formula
        ws[f'E{row}'] = "→"  # Trend placeholder
        ws[f'F{row}'] = ""   # Notes
        row += 1
    
    # Gap Summary Section
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "GAP SUMMARY BY RISK LEVEL"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    gap_headers = ['Risk Level', 'Count', '% of Total', '', '', '']
    for col_idx, header in enumerate(gap_headers, start=1):
        if header:
            cell = ws.cell(row=row, column=col_idx, value=header)
            apply_style(cell, styles["column_header"])
    row += 1
    
    gap_levels = [
        ("Critical", '=COUNTIF(Gap_Analysis!D$4:D$43,"Critical")', '=B{}/B16*100'),
        ("High", '=COUNTIF(Gap_Analysis!D$4:D$43,"High")', '=B{}/B16*100'),
        ("Medium", '=COUNTIF(Gap_Analysis!D$4:D$43,"Medium")', '=B{}/B16*100'),
        ("Low", '=COUNTIF(Gap_Analysis!D$4:D$43,"Low")', '=B{}/B16*100'),
    ]
    
    for level, count_formula, pct_formula in gap_levels:
        ws[f'A{row}'] = level
        ws[f'B{row}'] = count_formula
        ws[f'C{row}'] = pct_formula.format(row)
        row += 1
    
    # Evidence Completeness Section
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "EVIDENCE COMPLETENESS"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    evidence_headers = ['Metric', 'Value', '', '', '', '']
    for col_idx, header in enumerate(evidence_headers, start=1):
        if header:
            cell = ws.cell(row=row, column=col_idx, value=header)
            apply_style(cell, styles["column_header"])
    row += 1
    
    evidence_metrics = [
        ("Evidence Items Collected", '=COUNTA(Evidence_Register!A$4:A$105)'),
        ("Evidence Items Verified", '=COUNTIF(Evidence_Register!H$4:H$105,"Verified")'),
        ("Evidence Completeness %", f'=B{row+1}/B{row}*100'),
    ]
    
    for metric, formula in evidence_metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = formula
        row += 1
    
    # Approval Section
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "APPROVAL SIGN-OFF"
    apply_style(ws[f'A{row}'], styles["subheader"])
    row += 1
    
    approval_headers = ['Approver', 'Name', 'Signature', 'Date', 'Status', '']
    for col_idx, header in enumerate(approval_headers, start=1):
        if header:
            cell = ws.cell(row=row, column=col_idx, value=header)
            apply_style(cell, styles["column_header"])
    row += 1
    
    approvers = ["DPO", "CISO", "Legal Counsel"]
    for approver in approvers:
        ws[f'A{row}'] = approver
        ws[f'B{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
        ws[f'C{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
        ws[f'D{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
        ws[f'E{row}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
        row += 1
    
    # Add approval status validation
    validations = create_data_validations()
    ws.add_data_validation(validations['approval_status'])
    for r in range(row-3, row):
        validations['approval_status'].add(ws[f'E{r}'])
    
    # Add DPO approval validation to E3
    ws.add_data_validation(validations['approval_status'])
    validations['approval_status'].add(ws['E3'])
    
    # Set column widths
    ws.column_dimensions['A'].width = WIDTH_EXTRA_WIDE
    ws.column_dimensions['B'].width = WIDTH_MEDIUM
    ws.column_dimensions['C'].width = WIDTH_MEDIUM
    ws.column_dimensions['D'].width = WIDTH_MEDIUM
    ws.column_dimensions['E'].width = WIDTH_NARROW
    ws.column_dimensions['F'].width = WIDTH_WIDE
    
    ws.freeze_panes = 'A8'


# ============================================================================
# SECTION 7: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    
    logger.info("=" * 78)
    logger.info(f"{WORKBOOK_ID} - {ASSESSMENT_AREA} Generator")
    logger.info(f"ISO/IEC 27001:2022 Control {CONTROL_ID}")
    logger.info("=" * 78)
    
    wb = create_workbook()
    styles = setup_styles()
    
    logger.info("\n[1/11] Creating Instructions & Legend...")
    create_instructions(wb["Instructions_Legend"], styles)
    
    logger.info("[2/11] Creating Classification Schema...")
    create_classification_schema(wb["Classification_Schema"], styles)
    
    logger.info("[3/11] Creating Sensitive Data Inventory...")
    create_sensitive_data_inventory(wb["Sensitive_Data_Inventory"], styles)
    
    logger.info("[4/11] Creating Data Location Mapping...")
    create_data_location_mapping(wb["Data_Location_Mapping"], styles)
    
    logger.info("[5/11] Creating Data Owner Assignment...")
    create_data_owner_assignment(wb["Data_Owner_Assignment"], styles)
    
    logger.info("[6/11] Creating Regulatory Mapping...")
    create_regulatory_mapping(wb["Regulatory_Mapping"], styles)
    
    logger.info("[7/11] Creating Labeling Methods...")
    create_labeling_methods(wb["Labeling_Methods"], styles)
    
    logger.info("[8/11] Creating Discovery Results...")
    create_discovery_results(wb["Discovery_Results"], styles)
    
    logger.info("[9/11] Creating Gap Analysis...")
    create_gap_analysis(wb["Gap_Analysis"], styles)
    
    logger.info("[10/11] Creating Evidence Register...")
    create_evidence_register(wb["Evidence_Register"], styles)
    
    logger.info("[11/11] Creating Summary Dashboard...")
    create_summary_dashboard(wb["Summary_Dashboard"], styles)
    
    # Save workbook
    filename = f"ISMS-IMP-A.8.12.2_Data_Classification_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    logger.info(f"\n\u2705 SUCCESS: {filename}")
    logger.info("\nWorkbook Structure:")
    logger.info("  \u2022 Instructions & Legend")
    logger.info("  \u2022 Classification Schema (4 levels + 6 custom)")
    logger.info("  \u2022 Sensitive Data Inventory (30 rows: 10 examples + 20 blank)")
    logger.info("  \u2022 Data Location Mapping (40 rows: 5 examples + 35 blank)")
    logger.info("  \u2022 Data Owner Assignment (30 rows: 3 examples + 27 blank)")
    logger.info("  \u2022 Regulatory Mapping (25 rows: 5 examples + 20 blank)")
    logger.info("  \u2022 Labeling Methods (20 rows: 4 examples + 16 blank)")
    logger.info("  \u2022 Discovery Results (30 rows: 3 examples + 27 blank)")
    logger.info("  \u2022 Gap Analysis (40 rows: 3 examples + 37 blank)")
    logger.info("  \u2022 Evidence Register (100 rows)")
    logger.info("  \u2022 Summary Dashboard (11 KPIs + compliance tracking)")
    logger.info("\n" + "=" * 78)
    logger.info("NEXT STEPS:")
    logger.info("1. Open the workbook in Excel/LibreOffice")
    logger.info("2. Document your classification schema (min 4 levels)")
    logger.info("3. Inventory all sensitive data categories")
    logger.info("4. Map data to storage locations")
    logger.info("5. Assign data owners and stewards")
    logger.info("6. Complete regulatory compliance mapping")
    logger.info("7. Gather evidence for all assessments")
    logger.info("8. Review Summary Dashboard for compliance score")
    logger.info("9. Obtain DPO/CISO/Legal approval")
    logger.info("=" * 78)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
