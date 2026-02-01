#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS A.8.12 - Assessment File Normalization Utility
================================================================================

ISO/IEC 27001:2022 Annex A Control A.8.12: Data Leakage Prevention
Quality Assurance Utility: File Naming Normalization

--------------------------------------------------------------------------------
PURPOSE
--------------------------------------------------------------------------------

Normalizes DLP assessment workbook filenames to exact format required for
dashboard external workbook formulas, ensuring reliable data consolidation
and audit evidence traceability.

This utility is a CRITICAL prerequisite step before generating the compliance
dashboard. It creates standardized filenames without date suffixes, allowing
external formulas to reference source assessment files consistently.

Without normalization: Dashboard external formulas fail (#REF! errors)
With normalization: Dashboard automatically pulls data from assessments

--------------------------------------------------------------------------------
WHAT IT DOES
--------------------------------------------------------------------------------

1. Scans current directory for completed DLP assessment workbooks
2. Validates document IDs in Instructions_Legend sheet
3. Copies files to normalized filenames (removes date/version suffixes)
4. Creates audit manifest documenting filename changes
5. Verifies all 4 required assessment files are present

Filename Transformation Examples:
    ISMS-IMP-A.8.12.1_DLP_Infrastructure_20250125.xlsx
    → ISMS-IMP-A.8.12.1.xlsx
    
    ISMS-IMP-A.8.12.2_Data_Classification_20250124_v2.xlsx
    → ISMS-IMP-A.8.12.2.xlsx
    
    ISMS-IMP-A.8.12.3_Channel_Coverage_FINAL.xlsx
    → ISMS-IMP-A.8.12.3.xlsx
    
    ISMS-IMP-A.8.12.4_Monitoring_Response_20250125.xlsx
    → ISMS-IMP-A.8.12.4.xlsx

Normalized Target Filenames:
    ISMS-IMP-A.8.12.1.xlsx (DLP Infrastructure Assessment)
    ISMS-IMP-A.8.12.2.xlsx (Data Classification Assessment)
    ISMS-IMP-A.8.12.3.xlsx (Channel Coverage Assessment)
    ISMS-IMP-A.8.12.4.xlsx (Monitoring & Response Assessment)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System:
    Python 3.8 or higher

Dependencies:
    openpyxl - Excel file reading for document ID validation
    
Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic:
    python3 normalize_assessment_files_a812.py

The script operates on the current working directory.
Place all completed assessment workbooks in the same directory before running.

Options:
    No command-line options required.
    Script automatically detects and processes assessment files.

Output:
    • Normalized copies of assessment files (ISMS-IMP-A.8.12.X.xlsx)
    • Audit manifest: A812_Normalization_Manifest_YYYYMMDD.txt
    • Console summary showing files processed

Examples:
    # Navigate to directory with completed assessments
    cd /path/to/DLP_Assessments/
    
    # Run normalization
    python3 normalize_assessment_files_a812.py
    
    # Verify normalized files created
    ls -l ISMS-IMP-A.8.12.*.xlsx

Post-Normalization:
    After normalization completes:
    1. Verify all 4 normalized files present (ISMS-IMP-A.8.12.1-4.xlsx)
    2. Generate dashboard: python3 generate_a812_5_compliance_dashboard.py
    3. Place dashboard in same directory as normalized files
    4. Open dashboard and click "Update Links"

--------------------------------------------------------------------------------
WORKFLOW INTEGRATION
--------------------------------------------------------------------------------

Position in A.8.12 Framework:
    Quality assurance step between assessment completion and dashboard generation
    CRITICAL prerequisite for external workbook formula functionality

Typical Usage Pattern:
    Run this utility AFTER completing all 4 assessments and BEFORE generating
    the compliance dashboard. External formulas in dashboard reference normalized
    filenames created by this utility.

Integration Workflow:
    1. Generate assessment workbooks (scripts 1-4)
    2. Complete assessments (manual work)
    3. Run THIS UTILITY (normalize filenames)         ← YOU ARE HERE
    4. Generate dashboard (script 5)
    5. Consolidate data (optional - consolidate_a812_dashboard.py)
    6. Present to CISO/auditors

Related Scripts:
    - generate_a812_1_dlp_infrastructure.py (creates assessment 1)
    - generate_a812_2_data_classification.py (creates assessment 2)
    - generate_a812_3_channel_coverage.py (creates assessment 3)
    - generate_a812_4_monitoring_response.py (creates assessment 4)
    - generate_a812_5_compliance_dashboard.py (requires normalized files)
    - consolidate_a812_dashboard.py (alternative to external formulas)

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.12
Utility Type:         Quality Assurance / File Naming Normalization
Framework Version:    1.0
Script Version:       1.0
Date:                 25.01.2025
Author:               [Organization] ISMS Implementation Team

--------------------------------------------------------------------------------
NOTES
--------------------------------------------------------------------------------

Original Files Preserved:
    This utility COPIES files to normalized names.
    Original assessment files remain unchanged.
    No data loss risk - original files serve as backup.

File Validation:
    Script validates each workbook contains correct document ID in
    Instructions_Legend sheet before normalizing filename.
    Prevents normalization of incorrect or corrupted files.

Audit Manifest:
    Normalization manifest (A812_Normalization_Manifest_YYYYMMDD.txt) provides
    audit trail documenting:
    • Original filename → Normalized filename mapping
    • File size and modification timestamps
    • Document ID validation results
    • Normalization date and operator
    
    Retain manifest for audit evidence traceability.

Re-Running Normalization:
    Safe to re-run if assessment files updated.
    Existing normalized files will be overwritten.
    Always generates fresh audit manifest.

================================================================================
END OF HEADER - SCRIPT CODE FOLLOWS
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
from datetime import datetime
from pathlib import Path
import os
import shutil
import sys

# =============================================================================
# Third-Party Imports
# =============================================================================
try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed")
    print("Install with: sudo apt install python3-openpyxl")
    sys.exit(1)

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# ============================================================================
# CONFIGURATION
# ============================================================================

# Expected document IDs and titles from assessment workbooks
EXPECTED_DOCS = {
    "ISMS-IMP-A.8.12.1": {
        "title": "DLP Infrastructure Assessment",
        "normalized": "ISMS-IMP-A.8.12.1.xlsx"
    },
    "ISMS-IMP-A.8.12.2": {
        "title": "Data Classification Assessment",
        "normalized": "ISMS-IMP-A.8.12.2.xlsx"
    },
    "ISMS-IMP-A.8.12.3": {
        "title": "Channel Coverage Assessment",
        "normalized": "ISMS-IMP-A.8.12.3.xlsx"
    },
    "ISMS-IMP-A.8.12.4": {
        "title": "Monitoring & Response Assessment",
        "normalized": "ISMS-IMP-A.8.12.4.xlsx"
    },
}


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_workbook(filepath):
    """
    Validate workbook contains expected document ID in Instructions_Legend sheet.
    
    Args:
        filepath: Path to Excel workbook
    
    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        # Open workbook in read-only mode
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # Check for Instructions_Legend sheet (A.8.12 naming convention)
        if "Instructions_Legend" not in wb.sheetnames:
            wb.close()
            return (None, None)
        
        ws = wb["Instructions_Legend"]
        
        # Look for Document ID in column A (typically rows 4-20)
        # Format: "Document ID" label in column A, value in column B
        for row in range(3, 25):
            cell_label = ws.cell(row=row, column=1).value
            
            if cell_label and "Document ID" in str(cell_label):
                doc_id_value = ws.cell(row=row, column=2).value
                
                if doc_id_value:
                    doc_id = str(doc_id_value).strip()
                    
                    # Check if it's one of our expected documents
                    if doc_id in EXPECTED_DOCS:
                        title = EXPECTED_DOCS[doc_id]["title"]
                        wb.close()
                        return (doc_id, title)
        
        wb.close()
        return (None, None)
        
    except Exception as e:
        print(f"    \u26A0\uFE0F  Error reading file: {e}")
        return (None, None)


def get_file_info(filepath):
    """Get file metadata for manifest."""
    stat = filepath.stat()
    return {
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime),
        "created": datetime.fromtimestamp(stat.st_ctime),
    }


# ============================================================================
# SCANNING FUNCTIONS
# ============================================================================

def scan_directory(directory):
    """
    Scan directory for assessment workbooks.
    
    Args:
        directory: Path to scan
    
    Returns:
        dict: {doc_id: {'path': Path, 'info': metadata}}
    """
    found_assessments = {}
    directory = Path(directory)
    
    # Get all .xlsx files, excluding temporary files
    xlsx_files = [f for f in directory.glob("*.xlsx") if not f.name.startswith("~$")]
    
    if not xlsx_files:
        print(f"\n\u274C No Excel files found in {directory}\n")
        return found_assessments
    
    print(f"\n🔍 Scanning {len(xlsx_files)} Excel file(s) in {directory}...\n")
    
    for filepath in sorted(xlsx_files):
        print(f"  Checking: {filepath.name}")
        doc_id, title = validate_workbook(filepath)
        
        if doc_id:
            print(f"    \u2705 Valid: {doc_id} - {title}")
            
            # Check for duplicates
            if doc_id in found_assessments:
                print(f"\n    \u26A0\uFE0F  WARNING: DUPLICATE FOUND FOR {doc_id}")
                print(f"        Previous: {found_assessments[doc_id]['path'].name}")
                print(f"        Current:  {filepath.name}\n")
                
                response = input("        Use CURRENT file? (y/n): ").strip().lower()
                
                if response == 'y':
                    found_assessments[doc_id] = {
                        'path': filepath,
                        'info': get_file_info(filepath)
                    }
                    print(f"        → Replaced with current file\n")
                else:
                    print(f"        → Keeping previous file\n")
            else:
                found_assessments[doc_id] = {
                    'path': filepath,
                    'info': get_file_info(filepath)
                }
        else:
            print(f"    ⏭️  Skipped (not a valid assessment workbook)")
        
        print()  # Blank line for readability
    
    return found_assessments


# ============================================================================
# MANIFEST CREATION
# ============================================================================

def create_manifest(output_dir, mapping, source_dir):
    """
    Create audit manifest documenting the normalization process.
    
    Args:
        output_dir: Destination directory
        mapping: Dictionary of normalized files
        source_dir: Source directory
    
    Returns:
        Path: Path to created manifest file
    """
    manifest_path = Path(output_dir) / "normalization_manifest_a812.txt"
    
    with open(manifest_path, 'w', encoding='utf-8') as f:
        # Header
        f.write("=" * 80 + "\n")
        f.write("ISMS ASSESSMENT FILE NORMALIZATION MANIFEST\n")
        f.write("ISO/IEC 27001:2022 - Control A.8.12: Data Leakage Prevention\n")
        f.write("=" * 80 + "\n\n")
        
        # Normalization metadata
        f.write("NORMALIZATION METADATA\n")
        f.write("-" * 80 + "\n")
        f.write(f"Normalization Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Source Directory:        {source_dir.resolve()}\n")
        f.write(f"Output Directory:        {output_dir.resolve()}\n")
        f.write(f"Files Normalized:        {len(mapping)}/4 required\n")
        f.write(f"Normalization Status:    {'COMPLETE' if len(mapping) == 4 else 'INCOMPLETE'}\n")
        f.write("\n")
        
        # File mapping details
        f.write("FILE MAPPING\n")
        f.write("-" * 80 + "\n\n")
        
        for doc_id in sorted(EXPECTED_DOCS.keys()):
            if doc_id in mapping:
                info = mapping[doc_id]
                source_path = info['path']
                normalized_name = info['normalized']
                file_info = info['info']
                
                f.write(f"Document ID:     {doc_id}\n")
                f.write(f"Title:           {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"Source File:     {source_path.name}\n")
                f.write(f"Normalized File: {normalized_name}\n")
                f.write(f"File Size:       {file_info['size']:,} bytes\n")
                f.write(f"Last Modified:   {file_info['modified'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Status:          \u2705 NORMALIZED\n")
                f.write("\n")
            else:
                f.write(f"Document ID:     {doc_id}\n")
                f.write(f"Title:           {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"Status:          \u274C NOT FOUND\n")
                f.write("\n")
        
        # Dashboard integration notes
        f.write("=" * 80 + "\n")
        f.write("DASHBOARD INTEGRATION NOTES\n")
        f.write("=" * 80 + "\n\n")
        f.write("Purpose:\n")
        f.write("  This normalization process enables the compliance dashboard to reference\n")
        f.write("  assessment data using standardized filenames, eliminating manual Find &\n")
        f.write("  Replace operations.\n\n")
        f.write("Normalized Filename Convention:\n")
        f.write("  - Domain 1: ISMS-IMP-A.8.12.1.xlsx (DLP Infrastructure)\n")
        f.write("  - Domain 2: ISMS-IMP-A.8.12.2.xlsx (Data Classification)\n")
        f.write("  - Domain 3: ISMS-IMP-A.8.12.3.xlsx (Channel Coverage)\n")
        f.write("  - Domain 4: ISMS-IMP-A.8.12.4.xlsx (Monitoring & Response)\n\n")
        f.write("Dashboard Setup:\n")
        f.write("  1. Generate dashboard workbook:\n")
        f.write("     python3 generate_a812_5_compliance_dashboard.py\n\n")
        f.write("  2. Place dashboard in same directory as normalized files:\n")
        f.write(f"     {output_dir.resolve()}\n\n")
        f.write("  3. Open dashboard and click 'Update Links' when prompted\n\n")
        f.write("  4. Dashboard formulas auto-populate with current data!\n\n")
        f.write("External Link Behavior:\n")
        f.write("  - Dashboard uses Excel external link formulas\n")
        f.write("  - External workbook links auto-update when sources change\n")
        f.write("  - Place dashboard in same directory as normalized files\n")
        f.write("  - Open dashboard and click 'Update Links' when prompted\n\n")
        f.write("File Retention:\n")
        f.write("  - Original source files: Retained in source directory\n")
        f.write("  - Normalized copies: Located in output directory\n")
        f.write("  - This manifest: Retained with normalized files for audit\n\n")
        
        f.write("=" * 80 + "\n")
        f.write("END OF MANIFEST\n")
        f.write("=" * 80 + "\n")
    
    return manifest_path


# ============================================================================
# MAIN NORMALIZATION FUNCTION
# ============================================================================

def normalize_files(source_dir=None, output_dir=None, auto_confirm=False):
    """
    Main normalization orchestration function.
    
    Args:
        source_dir: Source directory (prompts if None)
        output_dir: Output directory (prompts if None)
        auto_confirm: Skip confirmation prompts (for automation)
    
    Returns:
        bool: True if successful, False otherwise
    """
    print("=" * 80)
    print("ISMS ASSESSMENT FILE NORMALIZATION UTILITY")
    print("ISO/IEC 27001:2022 - Control A.8.12: Data Leakage Prevention")
    print("=" * 80)
    print("\nThis script prepares DLP assessment workbooks for dashboard linking by:")
    print("  1. Scanning for completed assessment files")
    print("  2. Validating document IDs in Instructions_Legend sheet")
    print("  3. Copying to normalized filenames (no dates/versions)")
    print("  4. Creating audit manifest for traceability\n")
    
    # Get source directory
    if not source_dir:
        source_dir = input("Enter source directory path (or press Enter for current directory): ").strip()
        if not source_dir:
            source_dir = "."
    
    source_dir = Path(source_dir).resolve()
    
    if not source_dir.exists():
        print(f"\n\u274C Error: Source directory does not exist: {source_dir}\n")
        return False
    
    if not source_dir.is_dir():
        print(f"\n\u274C Error: Not a directory: {source_dir}\n")
        return False
    
    print(f"\n📁 Source directory: {source_dir}")
    
    # Get output directory
    if not output_dir:
        default_output = source_dir / "Dashboard_Sources"
        output_input = input(f"\nEnter output directory (or press Enter for '{default_output.name}'): ").strip()
        if not output_input:
            output_dir = default_output
        else:
            output_dir = Path(output_input)
    
    output_dir = Path(output_dir).resolve()
    
    # Create output directory if it doesn't exist
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        print(f"📁 Output directory: {output_dir}")
    except Exception as e:
        print(f"\n\u274C Error creating output directory: {e}\n")
        return False
    
    # Scan for assessment files
    found = scan_directory(source_dir)
    
    if not found:
        print("\u274C No valid assessment workbooks found in source directory\n")
        print("   Ensure files contain valid Document IDs in 'Instructions_Legend' sheet:")
        for doc_id, info in EXPECTED_DOCS.items():
            print(f"   - {doc_id}: {info['title']}")
        print()
        return False
    
    # Show normalization summary
    print("=" * 80)
    print("NORMALIZATION SUMMARY")
    print("=" * 80 + "\n")
    print(f"Found {len(found)} of {len(EXPECTED_DOCS)} required assessment workbooks:\n")
    
    mapping = {}
    for doc_id in sorted(EXPECTED_DOCS.keys()):
        if doc_id in found:
            normalized_name = EXPECTED_DOCS[doc_id]["normalized"]
            source_path = found[doc_id]['path']
            
            print(f"  \u2705 {doc_id}")
            print(f"     Title:      {EXPECTED_DOCS[doc_id]['title']}")
            print(f"     Source:     {source_path.name}")
            print(f"     Normalized: {normalized_name}\n")
            
            mapping[doc_id] = {
                'path': source_path,
                'normalized': normalized_name,
                'info': found[doc_id]['info']
            }
        else:
            print(f"  \u274C {doc_id} - NOT FOUND")
            print(f"     Title: {EXPECTED_DOCS[doc_id]['title']}\n")
    
    if len(found) < len(EXPECTED_DOCS):
        print(f"\u26A0\uFE0F  WARNING: Only {len(found)}/{len(EXPECTED_DOCS)} assessment workbooks found")
        print("   Dashboard will have incomplete data for missing assessments\n")
    
    print(f"Output directory: {output_dir}\n")
    
    # Confirm normalization
    if not auto_confirm:
        response = input("Proceed with normalization? (y/n): ").strip().lower()
        if response != 'y':
            print("\n\u274C Normalization cancelled by user\n")
            return False
    
    # Perform normalization (copy files)
    print("\n" + "=" * 80)
    print("NORMALIZING FILES...")
    print("=" * 80 + "\n")
    
    for doc_id, info in mapping.items():
        source = info['path']
        dest = output_dir / info['normalized']
        
        print(f"Copying: {source.name}")
        print(f"      → {dest.name}")
        
        try:
            shutil.copy2(source, dest)
            print(f"      \u2705 Success\n")
        except Exception as e:
            print(f"      \u274C Error: {e}\n")
            print(f"\u274C Normalization failed at {doc_id}\n")
            return False
    
    # Create audit manifest
    print("📄 Creating audit manifest...")
    try:
        manifest = create_manifest(output_dir, mapping, source_dir)
        print(f"   \u2705 Created: {manifest.name}\n")
    except Exception as e:
        print(f"   \u274C Error creating manifest: {e}\n")
        return False
    
    # Success summary
    print("=" * 80)
    print("\u2705 NORMALIZATION COMPLETE")
    print("=" * 80 + "\n")
    print(f"Normalized files:  {output_dir}")
    print(f"Audit manifest:    {manifest}\n")
    print("NEXT STEPS:\n")
    print("  1. Review audit manifest for file mapping details")
    print("  2. Generate dashboard workbook:")
    print("     python3 generate_a812_5_compliance_dashboard.py\n")
    print("  3. Place dashboard in same directory as normalized files:")
    print(f"     {output_dir}\n")
    print("  4. Open dashboard and click 'Update Links' when prompted\n")
    print("  5. Dashboard will auto-populate with current compliance data\n")
    print("=" * 80 + "\n")
    
    return True


# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Normalize ISMS DLP assessment workbooks for dashboard integration",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode (prompts for directories)
  python3 normalize_assessment_files_a812.py
  
  # Specify source directory
  python3 normalize_assessment_files_a812.py --source ./assessments
  
  # Specify both directories
  python3 normalize_assessment_files_a812.py --source ./assessments --output ./dashboard
  
  # Automated mode (no prompts)
  python3 normalize_assessment_files_a812.py --source ./assessments --auto-confirm
        """
    )
    
    parser.add_argument(
        '--source', '-s',
        help='Source directory containing assessment workbooks (default: current directory)'
    )
    
    parser.add_argument(
        '--output', '-o',
        help='Output directory for normalized files (default: ./Dashboard_Sources)'
    )
    
    parser.add_argument(
        '--auto-confirm', '-y',
        action='store_true',
        help='Skip confirmation prompts (for automation)'
    )
    
    args = parser.parse_args()
    
    success = normalize_files(
        source_dir=args.source,
        output_dir=args.output,
        auto_confirm=args.auto_confirm
    )
    
    sys.exit(0 if success else 1)


# ============================================================================
# END OF NORMALIZATION SCRIPT
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION (syntax validated, structure verified)
# QA_TOOL: Claude Code Standardization
# =============================================================================
