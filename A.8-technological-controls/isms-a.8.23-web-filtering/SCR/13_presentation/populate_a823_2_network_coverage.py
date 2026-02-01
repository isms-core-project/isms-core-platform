#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
ISMS-IMP-A.8.23.2 - Network Coverage Assessment
Comprehensive Data Population Script for CISO Presentation

Populates ALL sheets with production-quality demo data:
- Network Segment Inventory (15+ segments)
- Coverage Matrix (segment vs. filtering solution)
- Gap Identification (uncovered segments)
- Device Inventory (25+ device categories)
- Exemption Register (documented bypasses)
- Coverage Verification (test results)
- Evidence Register (18+ documents)
- Approval Sign-Off (complete workflow)

Usage:
    python3 populate_a823_2_network_coverage.py ISMS-IMP-A.8.23.2.xlsx

Generated: 2026-02-01
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


def safely_write_data(ws, start_row, data):
    """Safely write data to worksheet, handling merged cells."""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except Exception as e:  # Skip merged/protected cells
                continue
        entries_written += 1
    return entries_written


def populate_network_segment_inventory(wb):
    """Populate Network Segment Inventory."""
    ws = wb["Network_Segment_Inventory"]

    data = [
        # Segment Name, VLAN/Subnet, Location, Classification, Users/Devices, Purpose, Owner, Filter Status
        ["Corporate LAN - Zurich HQ", "VLAN 10 / 10.10.0.0/16", "Zurich Office", "Confidential", "1,200 users", "Primary corporate network", "IT Infrastructure", "Covered - Zscaler"],
        ["Corporate LAN - Geneva", "VLAN 11 / 10.11.0.0/16", "Geneva Office", "Confidential", "450 users", "Branch office network", "IT Infrastructure", "Covered - Zscaler"],
        ["Corporate LAN - Basel", "VLAN 12 / 10.12.0.0/16", "Basel Office", "Confidential", "280 users", "Branch office network", "IT Infrastructure", "Covered - Zscaler"],
        ["Remote Workers (VPN)", "Dynamic / 10.200.0.0/16", "Remote", "Confidential", "2,500 users", "VPN-connected remote workers", "IT Security", "Covered - Zscaler ZCC"],
        ["Remote Workers (Direct)", "Dynamic / Public IPs", "Remote", "Confidential", "500 users", "Zscaler Client direct connect", "IT Security", "Covered - Zscaler ZCC"],
        ["Development Network", "VLAN 50 / 10.50.0.0/24", "Zurich DC", "Restricted", "150 developers", "Development environment", "Development", "Covered - Zscaler"],
        ["Production Servers", "VLAN 100 / 10.100.0.0/24", "Zurich DC", "Highly Restricted", "500 servers", "Production infrastructure", "IT Operations", "Covered - Zscaler Cloud Connector"],
        ["DMZ - Web Servers", "VLAN 200 / 172.16.0.0/24", "Zurich DC", "Public", "25 servers", "Public-facing web servers", "Web Operations", "Covered - Inline inspection"],
        ["AWS VPC - Production", "VPC / 10.150.0.0/16", "AWS eu-central-1", "Highly Restricted", "1,000 instances", "Cloud production workloads", "Cloud Team", "Covered - Zscaler Cloud Connector"],
        ["Azure VNet - Production", "VNet / 10.160.0.0/16", "Azure West Europe", "Highly Restricted", "500 instances", "Cloud production workloads", "Cloud Team", "Partial - Migration in progress"],
        ["GCP VPC - Analytics", "VPC / 10.170.0.0/16", "GCP europe-west6", "Restricted", "200 instances", "Data analytics workloads", "Data Team", "Covered - Prisma Access"],
        ["Guest WiFi", "VLAN 250 / 192.168.250.0/24", "All Offices", "Public", "Variable", "Guest and visitor access", "Facilities", "Partial - DNS only (Umbrella)"],
        ["IoT Network", "VLAN 300 / 10.30.0.0/24", "All Offices", "Internal", "500 devices", "IoT sensors and devices", "Facilities", "❌ Non-Compliant - Legacy proxy"],
        ["SCADA/OT Network", "VLAN 400 / 10.40.0.0/24", "Factory Floor", "Highly Restricted", "100 devices", "Industrial control systems", "Manufacturing", "Exempt - Air-gapped"],
        ["Partner Extranet", "VLAN 500 / 172.20.0.0/24", "Zurich DC", "Restricted", "50 partners", "Partner access zone", "Partner Management", "Covered - Zscaler ZPA"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Network_Segment_Inventory: {count} segments documented")
    return count


def populate_coverage_matrix(wb):
    """Populate Coverage Matrix."""
    ws = wb["Coverage_Matrix"]

    data = [
        # Segment, Zscaler ZIA, Zscaler ZPA, Prisma Access, Cisco Umbrella, Legacy Proxy, Coverage %
        ["Corporate LAN - Zurich HQ", "Active", "-", "-", "-", "-", "100%"],
        ["Corporate LAN - Geneva", "Active", "-", "-", "-", "-", "100%"],
        ["Corporate LAN - Basel", "Active", "-", "-", "-", "-", "100%"],
        ["Remote Workers (VPN)", "Active (ZCC)", "-", "-", "-", "-", "100%"],
        ["Remote Workers (Direct)", "Active (ZCC)", "-", "-", "-", "-", "100%"],
        ["Development Network", "Active", "-", "-", "-", "-", "100%"],
        ["Production Servers", "Active (CC)", "-", "-", "-", "-", "100%"],
        ["DMZ - Web Servers", "Active", "-", "-", "-", "-", "100%"],
        ["AWS VPC - Production", "Active (CC)", "-", "-", "-", "-", "100%"],
        ["Azure VNet - Production", "⚠️ Partial", "-", "Active", "-", "-", "85%"],
        ["GCP VPC - Analytics", "-", "-", "Active", "-", "-", "100%"],
        ["Guest WiFi", "-", "-", "-", "DNS Only", "-", "40%"],
        ["IoT Network", "-", "-", "-", "-", "Squid", "25%"],
        ["SCADA/OT Network", "-", "-", "-", "-", "-", "Exempt"],
        ["Partner Extranet", "-", "Active", "-", "-", "-", "100%"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Coverage_Matrix: {count} segments mapped")
    return count


def populate_gap_identification(wb):
    """Populate Gap Identification."""
    ws = wb["Gap_Identification"]

    data = [
        # Gap ID, Segment, Current Coverage, Gap Description, Risk Level, Business Impact, Remediation Plan, Target Date
        ["COV-GAP-001", "Azure VNet - Production", "85% (Prisma)", "15% of Azure workloads not covered", "High", "Cloud workloads may access malicious content", "Extend Zscaler Cloud Connector to remaining Azure subnets", "2026-04-30"],
        ["COV-GAP-002", "Guest WiFi", "40% (DNS only)", "No inline content inspection for guests", "Medium", "Guests could access malicious downloads", "Implement captive portal with Zscaler redirect", "2026-05-31"],
        ["COV-GAP-003", "IoT Network", "25% (Legacy Squid)", "IoT devices use outdated proxy without SSL inspection", "High", "IoT devices vulnerable to C2 and malware", "Deploy Zscaler IoT solution, retire Squid", "2026-06-30"],
        ["COV-GAP-004", "Mobile Devices (BYOD)", "Variable", "Personal devices not always using Zscaler Client", "Medium", "BYOD accessing corporate data without filtering", "Enforce ZCC via MDM conditional access", "2026-03-31"],
        ["COV-GAP-005", "Contractor Laptops", "60%", "Some contractors bypass filtering", "Medium", "Contractors may introduce threats", "Mandate ZCC for all contractor devices", "2026-04-15"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Gap_Identification: {count} gaps identified")
    return count


def populate_device_inventory(wb):
    """Populate Device Inventory."""
    ws = wb["Device_Inventory"]

    data = [
        # Device Type, Count, Network Segment, Filtering Method, Coverage Status, MDM Managed, Notes
        ["Windows Workstations", "2,800", "Corporate LAN", "Zscaler Client Connector", "Covered", "Yes - Intune", "Auto-enrolled via MDM"],
        ["Windows Laptops", "1,500", "Remote/Corporate", "Zscaler Client Connector", "Covered", "Yes - Intune", "Always-on Z-Tunnel 2.0"],
        ["MacBooks", "800", "Remote/Corporate", "Zscaler Client Connector", "Covered", "Yes - Jamf", "macOS 12+ required"],
        ["Linux Workstations (Dev)", "150", "Development Network", "PAC file + explicit proxy", "Covered", "No - Manual", "Developer exception process"],
        ["Windows Servers", "400", "Production/DMZ", "Cloud Connector", "Covered", "No", "Server 2019+ only"],
        ["Linux Servers", "600", "Production/Cloud", "Cloud Connector", "Covered", "No", "Ubuntu 20.04+ / RHEL 8+"],
        ["iOS Devices (Corporate)", "1,200", "Remote/Corporate", "Zscaler Client Connector", "Covered", "Yes - Intune", "Supervised mode required"],
        ["Android Devices (Corporate)", "400", "Remote/Corporate", "Zscaler Client Connector", "Covered", "Yes - Intune", "Android Enterprise enrolled"],
        ["iOS Devices (BYOD)", "800", "Remote", "Zscaler Client Connector", "⚠️ Partial", "Partial - MAM", "App-level protection only"],
        ["Android Devices (BYOD)", "300", "Remote", "Zscaler Client Connector", "⚠️ Partial", "Partial - MAM", "App-level protection only"],
        ["Printers (Network)", "150", "Corporate LAN", "N/A", "Exempt", "No", "No internet access required"],
        ["Video Conferencing Systems", "50", "Corporate LAN", "PAC file", "Covered", "No", "Poly/Cisco systems"],
        ["IP Phones (VoIP)", "2,000", "Voice VLAN", "N/A", "Exempt", "No", "Isolated VLAN, no data access"],
        ["Security Cameras", "200", "IoT Network", "Legacy proxy (Squid)", "❌ Non-Compliant", "No", "Migration planned Q2"],
        ["Building Access Controllers", "75", "IoT Network", "Legacy proxy (Squid)", "❌ Non-Compliant", "No", "Migration planned Q2"],
        ["Environmental Sensors", "100", "IoT Network", "Legacy proxy (Squid)", "❌ Non-Compliant", "No", "Migration planned Q2"],
        ["Smart Displays", "80", "IoT Network", "Legacy proxy (Squid)", "❌ Non-Compliant", "No", "Migration planned Q2"],
        ["Guest Devices (Variable)", "~200/day", "Guest WiFi", "Cisco Umbrella (DNS)", "⚠️ Partial", "No", "Captive portal planned"],
        ["Contractor Laptops", "200", "Corporate/Remote", "Mixed", "⚠️ Partial", "Partial", "Enforcement in progress"],
        ["Kiosk Devices", "25", "Corporate LAN", "Zscaler + Kiosk mode", "Covered", "Yes - Intune", "Locked down configuration"],
        ["AWS EC2 Instances", "800", "AWS VPC", "Cloud Connector", "Covered", "N/A", "Auto-scaled workloads"],
        ["Azure VMs", "400", "Azure VNet", "Prisma Access / Cloud Connector", "⚠️ Partial", "N/A", "Migration in progress"],
        ["GCP Compute Instances", "200", "GCP VPC", "Prisma Access", "Covered", "N/A", "Analytics workloads"],
        ["Kubernetes Pods (AWS EKS)", "1,500", "AWS VPC", "Cloud Connector", "Covered", "N/A", "Egress via node proxy"],
        ["Docker Containers (On-Prem)", "300", "Production Servers", "Host-based routing", "Covered", "N/A", "Routed through Zscaler"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Device_Inventory: {count} device categories documented")
    return count


def populate_exemption_register(wb):
    """Populate Exemption Register."""
    ws = wb["Exemption_Register"]

    data = [
        # Exemption ID, System/Segment, Reason, Compensating Control, Risk Acceptance, Approved By, Expiry Date, Status
        ["EXM-001", "SCADA/OT Network", "Air-gapped network, industrial safety requirement", "Network isolation, dedicated firewalls, no internet", "Accepted - Manufacturing requirement", "CISO", "2027-01-31", "Active"],
        ["EXM-002", "VoIP Network (VLAN)", "Voice quality requirements, no web traffic", "Isolated VLAN, ACLs block web access", "Accepted - Operational", "Network Security", "2026-12-31", "Active"],
        ["EXM-003", "Network Printers", "No browser/internet capability", "No route to internet, management VLAN isolated", "Accepted - Technical limitation", "IT Security", "2026-12-31", "Active"],
        ["EXM-004", "Video Conferencing (Poly)", "Vendor-specific traffic requirements", "PAC file for approved domains only, logging enabled", "Accepted - Business critical", "IT Security", "2026-06-30", "Active"],
        ["EXM-005", "Financial Trading Terminal", "Latency-sensitive Bloomberg/Reuters traffic", "Direct route for approved IPs, all other traffic filtered", "Accepted - Business critical", "CISO", "2026-06-30", "Active"],
        ["EXM-006", "Security Research Lab", "Malware analysis requires unrestricted access", "Isolated network, no corporate connectivity, monitored", "Accepted - Security operations", "CISO", "2026-12-31", "Active"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Exemption_Register: {count} exemptions documented")
    return count


def populate_coverage_verification(wb):
    """Populate Coverage Verification test results."""
    ws = wb["Coverage_Verification"]

    data = [
        # Test ID, Segment, Test Type, Test Date, Tester, Result, Evidence, Notes
        ["VER-001", "Corporate LAN - Zurich", "Malware download test", "2026-01-20", "Security Team", "PASS - Blocked", "Block log EV2-015", "EICAR test file blocked"],
        ["VER-002", "Corporate LAN - Zurich", "Phishing URL access", "2026-01-20", "Security Team", "PASS - Blocked", "Block log EV2-016", "Known phishing URL blocked"],
        ["VER-003", "Corporate LAN - Geneva", "Category block test", "2026-01-21", "Security Team", "PASS - Blocked", "Block log EV2-017", "Adult content category blocked"],
        ["VER-004", "Remote Workers (ZCC)", "SSL inspection verification", "2026-01-22", "Security Team", "PASS - Inspected", "SSL test log", "TLS 1.3 decrypted successfully"],
        ["VER-005", "AWS VPC - Production", "Cloud Connector test", "2026-01-23", "Cloud Team", "PASS - Filtered", "AWS flow logs", "EC2 egress filtered"],
        ["VER-006", "Azure VNet - Production", "Coverage gap test", "2026-01-23", "Cloud Team", "PARTIAL - 85%", "Azure NSG logs", "Subnet 10.160.50.0/24 not covered"],
        ["VER-007", "Guest WiFi", "Malware download test", "2026-01-24", "Security Team", "PARTIAL - DNS only", "Umbrella logs", "Blocked at DNS, no file inspection"],
        ["VER-008", "IoT Network", "C2 callback test", "2026-01-24", "Security Team", "FAIL - Not blocked", "Squid logs", "Legacy proxy lacks threat intel"],
        ["VER-009", "Development Network", "Bypass attempt test", "2026-01-25", "Red Team", "PASS - Blocked", "Block log EV2-018", "Proxy bypass attempt detected/blocked"],
        ["VER-010", "Partner Extranet (ZPA)", "Access control test", "2026-01-25", "Security Team", "PASS - Controlled", "ZPA access logs", "App-level access verified"],
        ["VER-011", "Mobile (iOS)", "ZCC tunnel test", "2026-01-26", "Mobile Team", "PASS - Tunneled", "ZCC logs", "Always-on tunnel verified"],
        ["VER-012", "Mobile (Android)", "Split tunnel test", "2026-01-26", "Mobile Team", "PASS - Configured", "ZCC logs", "Split tunnel for approved apps only"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Coverage_Verification: {count} test results documented")
    return count


def populate_evidence_register(wb):
    """Populate Evidence Register with audit evidence."""
    ws = wb["Evidence_Register"]

    data = [
        # Evidence ID, Title, Type, Related Sheet, Location, Collection Date, Owner, Retention, Status
        ["EV2-001", "Network Segment Inventory Export", "Inventory", "Network_Segment_Inventory", "SharePoint/Network/Inventory", "2026-01-28", "Network Team", "3 years", "Current"],
        ["EV2-002", "VLAN Configuration Export", "Technical Config", "Network_Segment_Inventory", "Git/network/vlan-configs", "2026-01-28", "Network Team", "Version controlled", "Current"],
        ["EV2-003", "Zscaler Location Configuration", "Technical Config", "Coverage_Matrix", "SharePoint/Evidence/Zscaler", "2026-01-25", "Security Team", "3 years", "Current"],
        ["EV2-004", "Cloud Connector Deployment Docs", "Architecture", "Coverage_Matrix", "SharePoint/Cloud/Zscaler-CC", "2026-01-20", "Cloud Team", "3 years", "Current"],
        ["EV2-005", "Gap Analysis Report Q1 2026", "Assessment", "Gap_Identification", "SharePoint/Security/GapAnalysis", "2026-01-28", "Security Team", "3 years", "Current"],
        ["EV2-006", "CMDB Device Export", "Inventory", "Device_Inventory", "ServiceNow/CMDB-Export", "2026-01-28", "IT Operations", "1 year", "Current"],
        ["EV2-007", "Intune Device Compliance Report", "Compliance", "Device_Inventory", "SharePoint/Endpoint/Intune", "2026-01-28", "Endpoint Team", "1 year", "Current"],
        ["EV2-008", "Jamf Mac Enrollment Report", "Compliance", "Device_Inventory", "SharePoint/Endpoint/Jamf", "2026-01-28", "Endpoint Team", "1 year", "Current"],
        ["EV2-009", "Exemption Approval Records", "Approval", "Exemption_Register", "SharePoint/Security/Exemptions", "2026-01-15", "IT Security", "5 years", "Current"],
        ["EV2-010", "Risk Acceptance Forms", "Risk", "Exemption_Register", "SharePoint/Risk/Acceptances", "2026-01-15", "Risk Management", "7 years", "Current"],
        ["EV2-011", "Coverage Test Plan", "Test Plan", "Coverage_Verification", "SharePoint/QA/TestPlans", "2026-01-15", "Security Team", "3 years", "Current"],
        ["EV2-012", "Coverage Test Results Summary", "Test Report", "Coverage_Verification", "SharePoint/QA/TestResults", "2026-01-26", "Security Team", "3 years", "Current"],
        ["EV2-013", "Penetration Test Report (Network)", "Audit Report", "Coverage_Verification", "SharePoint/Audits/PenTest2025", "2025-11-30", "External Auditor", "5 years", "Current"],
        ["EV2-014", "Zscaler Traffic Logs (Sample)", "Logs", "Coverage_Verification", "Splunk/zscaler-index", "2026-01-28", "SOC Team", "90 days", "Current"],
        ["EV2-015", "Malware Block Log (EICAR Test)", "Test Evidence", "Coverage_Verification", "SharePoint/Evidence/BlockTests", "2026-01-20", "Security Team", "1 year", "Current"],
        ["EV2-016", "Phishing Block Log", "Test Evidence", "Coverage_Verification", "SharePoint/Evidence/BlockTests", "2026-01-20", "Security Team", "1 year", "Current"],
        ["EV2-017", "Category Block Test Log", "Test Evidence", "Coverage_Verification", "SharePoint/Evidence/BlockTests", "2026-01-21", "Security Team", "1 year", "Current"],
        ["EV2-018", "Bypass Attempt Detection Log", "Security Event", "Coverage_Verification", "Splunk/security-alerts", "2026-01-25", "SOC Team", "1 year", "Current"],
    ]

    count = safely_write_data(ws, 6, data)
    logger.info(f"    Evidence_Register: {count} evidence documents")
    return count


def safe_cell_write(ws, cell_ref, value):
    """Safely write to a cell, handling merged cells."""
    try:
        cell = ws[cell_ref]
        if not isinstance(cell, MergedCell):
            cell.value = value
    except Exception as e:  # Skip merged/protected cells
        pass


def populate_approval_signoff(wb):
    """Populate Approval Sign-Off with stakeholder workflow."""
    ws = wb["Approval_Sign_Off"]

    # Assessor information
    safe_cell_write(ws, "B5", "Stefan Brunner")
    safe_cell_write(ws, "B6", "Network Security Engineer")
    safe_cell_write(ws, "B7", "stefan.brunner@company.com")
    safe_cell_write(ws, "B8", datetime.now().strftime("%d.%m.%Y"))

    # Technical reviewer
    safe_cell_write(ws, "B12", "Michael Keller")
    safe_cell_write(ws, "B13", "Network Architecture Lead")
    safe_cell_write(ws, "B14", "michael.keller@company.com")
    safe_cell_write(ws, "B15", (datetime.now() + timedelta(days=3)).strftime("%d.%m.%Y"))
    safe_cell_write(ws, "B16", "Verified - Coverage gaps accurately documented. IoT migration is critical priority.")

    # Security reviewer
    safe_cell_write(ws, "B20", "Anna Müller")
    safe_cell_write(ws, "B21", "Information Security Manager")
    safe_cell_write(ws, "B22", "anna.mueller@company.com")
    safe_cell_write(ws, "B23", (datetime.now() + timedelta(days=5)).strftime("%d.%m.%Y"))
    safe_cell_write(ws, "B24", "Approved - Exemptions properly documented with compensating controls.")

    # Management approval
    safe_cell_write(ws, "B28", "Dr. Klaus Fischer")
    safe_cell_write(ws, "B29", "CISO")
    safe_cell_write(ws, "B30", "klaus.fischer@company.com")
    safe_cell_write(ws, "B31", (datetime.now() + timedelta(days=7)).strftime("%d.%m.%Y"))
    safe_cell_write(ws, "B32", "Approved. Prioritize IoT and Guest WiFi coverage improvements.")

    # Next review
    safe_cell_write(ws, "B36", (datetime.now() + timedelta(days=90)).strftime("%d.%m.%Y"))
    safe_cell_write(ws, "B37", "Network Security Team")

    logger.info("    Approval_Sign_Off: Complete workflow populated")


def main():
    """Main function to populate the workbook."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.8.23.2 - Network Coverage Assessment")
    logger.info("Comprehensive Data Population for CISO Presentation")
    logger.info("=" * 80)

    if len(sys.argv) < 2:
        logger.error("Usage: python3 populate_a823_2_network_coverage.py <workbook.xlsx>")
        sys.exit(1)

    filepath = sys.argv[1]
    logger.info(f"Loading: {filepath}")

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        logger.error(f"Failed to load workbook: {e}")
        sys.exit(1)

    logger.info(f"Sheets found: {len(wb.sheetnames)}")

    total_entries = 0

    logger.info("[1/7] Populating Network Segment Inventory...")
    total_entries += populate_network_segment_inventory(wb)

    logger.info("[2/7] Populating Coverage Matrix...")
    total_entries += populate_coverage_matrix(wb)

    logger.info("[3/7] Populating Gap Identification...")
    total_entries += populate_gap_identification(wb)

    logger.info("[4/7] Populating Device Inventory...")
    total_entries += populate_device_inventory(wb)

    logger.info("[5/7] Populating Exemption Register...")
    total_entries += populate_exemption_register(wb)

    logger.info("[6/7] Populating Coverage Verification...")
    total_entries += populate_coverage_verification(wb)

    logger.info("[7/7] Populating Evidence Register & Approval...")
    evidence_count = populate_evidence_register(wb)
    populate_approval_signoff(wb)

    wb.save(filepath)
    logger.info(f"Saved: {filepath}")

    logger.info("=" * 80)
    logger.info("DATA POPULATION COMPLETE")
    logger.info("=" * 80)
    logger.info(f"Summary:")
    logger.info(f"  - Assessment Entries: {total_entries}")
    logger.info(f"  - Evidence Documents: {evidence_count}")
    logger.info(f"  - Total Data Points: {total_entries + evidence_count}")
    logger.info("CISO-Ready: Professional network coverage assessment")
    logger.info("=" * 80)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED - PRESENTATION MODE POPULATE SCRIPT
# QA_TOOL: Claude Code
# CHANGES: Created for CISO presentation demo data
# =============================================================================
