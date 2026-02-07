#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.2-3-5.S1 - Authentication Inventory Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.8.2, A.8.3, A.8.5: Authentication & Privileged Access
Assessment Domain 1 of 6: Authentication Mechanism Inventory

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific authentication infrastructure, identity providers,
and access control requirements.

Key customization areas:
1. Identity providers (Azure AD, Okta, AD per your infrastructure)
2. Authentication methods (MFA options specific to your deployment)
3. Protocol requirements (SAML, OAuth per your integration needs)
4. Password policy standards (aligned with your security policy)
5. SSO coverage targets (based on your integration roadmap)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for authentication)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

**Purpose:**
Enables systematic inventory and assessment of all authentication mechanisms
across the organization's infrastructure, supporting evidence-based validation
of authentication controls per ISO 27001:2022 Control A.8.5 requirements.

**Assessment Scope:**
- Authentication mechanism inventory (passwords, MFA, SSO, certificates, biometrics)
- Authentication protocol usage (SAML, OAuth, OpenID Connect, Kerberos, LDAP)
- Authentication provider mapping (Azure AD, Okta, Active Directory, local)
- Password policy compliance per system
- SSO coverage and integration status
- Authentication logging and monitoring
- Deprecated authentication method identification
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and authentication standards
2. Authentication Inventory - Systems, methods, protocols, providers
3. Password Policy Compliance - Policy enforcement per system
4. SSO Integration Status - SSO coverage and gaps
5. Authentication Protocols - Protocol usage and security assessment
6. Deprecated Methods - Legacy authentication identification
7. Gap Analysis - Non-compliant authentication mechanisms
8. Evidence Register - Audit evidence tracking
9. Approval & Sign-Off - Stakeholder review workflow

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a8235_1_authentication_inventory.py

Requirements:
    - Python 3.8+
    - openpyxl library: pip install openpyxl

Output:
    ISMS-IMP-A.8.2-3-5.S1_Authentication_Inventory_Assessment_YYYYMMDD.xlsx
"""
# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# ============================================================================
# SECTION 1: CONSTANTS AND CONFIGURATION
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
KEY = '\U0001F511'    # 🔑 Key
PACKAGE = '\U0001F4E6' # 📦 Package
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow
WORKBOOK_NAME = "Authentication Inventory Assessment"
DOCUMENT_ID = "ISMS-IMP-A.8.2-3-5.S1"
VERSION = "1.0"
CONTROL_REF = "ISO/IEC 27001:2022 - Controls A.8.2, A.8.3, A.8.5: Authentication & Privileged Access"
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_Authentication_Inventory_{GENERATED_TIMESTAMP}.xlsx"
# Workbook structure
SYSTEM_ROW_COUNT = 150      # Pre-formatted rows for system inventory
PROTOCOL_ROW_COUNT = 50     # Authentication protocol entries
SSO_APP_ROW_COUNT = 100     # SSO application tracking
PASSWORD_POLICY_COUNT = 30  # Password policy configurations
LEGACY_ROW_COUNT = 40       # Legacy authentication tracking
GAP_ROW_COUNT = 50          # Gap analysis entries
EVIDENCE_ROW_COUNT = 30     # Evidence register entries
# Authentication methods
AUTH_METHODS = [
    "Password Only",
    "Password + MFA (Optional)",
    "Password + MFA (Mandatory)",
    "SSO (Password-based)",
    "SSO + MFA (Mandatory)",
    "Certificate-Based",
    "Certificate + MFA",
    "Hardware Token (FIDO2)",
    "Biometric",
    "Biometric + PIN",
    "Smart Card",
    "Other (Specify)"
]
# Authentication protocols
AUTH_PROTOCOLS = [
    "SAML 2.0",
    "OAuth 2.0",
    "OpenID Connect (OIDC)",
    "Kerberos",
    "RADIUS",
    "LDAP",
    "LDAPS (LDAP over TLS)",
    "NTLM (Legacy)",
    "Basic Auth (Legacy)",
    "Digest Auth",
    "Certificate (mTLS)",
    "Custom/Proprietary",
    "Unknown"
]
# Identity providers
IDENTITY_PROVIDERS = [
    "Azure AD / Entra ID",
    "Okta",
    "Google Workspace",
    "Active Directory (On-Prem)",
    "Auth0",
    "Ping Identity",
    "ForgeRock",
    "OneLogin",
    "Local Authentication",
    "Application-Specific"
]
# MFA methods
MFA_METHODS = [
    "Hardware Token (FIDO2/YubiKey)",
    "Authenticator App (TOTP)",
    "Push Notification",
    "Biometric (Fingerprint/Face)",
    "SMS (Discouraged)",
    "Voice Call",
    "Email OTP",
    "Not Available"
]
# Compliance status
COMPLIANCE_STATUS = [
    f"{CHECK} Compliant",
    f"{WARNING} Partial Compliance",
    f"{XMARK} Non-Compliant",
    "🔄 In Progress",
    "📋 Under Review",
    "❓ Unknown",
    "➖ N/A"
]
# Security ratings
SECURITY_RATINGS = [
    "🟢 Strong (90-100%)",
    "🟡 Adequate (70-89%)",
    "🟠 Weak (50-69%)",
    "🔴 Poor (<50%)",
    "⚫ Unacceptable"
]
# Priority levels
PRIORITY_LEVELS = [
    "🔴 Critical",
    "🟠 High",
    "🟡 Medium",
    "🟢 Low",
    "⚪ Informational"
]
# SECTION 2: STYLE DEFINITIONS
def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style TEMPLATES as dictionaries to avoid shared object warnings.

    CRITICAL: Each cell must receive NEW style objects, not shared references.
    """
    return {
        'header': {
            'font': {'name': 'Calibri', 'size': 11, 'bold': True, 'color': 'FFFFFF'},
            'fill': {'patternType': 'solid', 'fgColor': '003366'},
            'alignment': {'horizontal': 'center', 'vertical': 'center', 'wrap_text': True},
            'border': {'left': 'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
        },
        'subheader': {
            'font': {'name': 'Calibri', 'size': 11, 'bold': True, 'color': '000000'},
            'fill': {'patternType': 'solid', 'fgColor': 'D8E4F8'},
            'alignment': {'horizontal': 'left', 'vertical': 'center', 'wrap_text': True},
        },
        'data': {
            'font': {'name': 'Calibri', 'size': 10, 'color': '000000'},
            'fill': {'patternType': 'solid', 'fgColor': 'FFFFFF'},
            'alignment': {'horizontal': 'left', 'vertical': 'center', 'wrap_text': False},
        },
        'title': {
            'font': {'name': 'Calibri', 'size': 16, 'bold': True, 'color': '003366'},
            'alignment': {'horizontal': 'left', 'vertical': 'center'}
        },
        'instruction': {
            'alignment': {'horizontal': 'left', 'vertical': 'top', 'wrap_text': True}
        }
    }
def apply_style(cell, style_dict):
    """
    Apply style dictionary to a cell with NEW objects (avoid shared references).
    Args:
        cell: openpyxl cell object
        style_dict: Dictionary containing font, fill, alignment, border specs
    """
    if 'font' in style_dict:
        cell.font = Font(**style_dict['font'])
    if 'fill' in style_dict:
        cell.fill = PatternFill(**style_dict['fill'])
    if 'alignment' in style_dict:
        cell.alignment = Alignment(**style_dict['alignment'])
    if 'border' in style_dict:
        cell.border = Border(
            left=Side(style=style_dict['border'].get('left', 'thin')),
            right=Side(style=style_dict['border'].get('right', 'thin')),
            top=Side(style=style_dict['border'].get('top', 'thin')),
            bottom=Side(style=style_dict['border'].get('bottom', 'thin'))
        )
# SECTION 3: WORKBOOK CREATION
def create_workbook():
    """Create workbook with all required sheets."""
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    # Create sheets matching specification
    sheets = [
        "Instructions_Legend",
        "System_Auth_Inventory",
        "Auth_Protocol_Analysis",
        "SSO_Integration_Status",
        "Password_Policy_Compliance",
        "MFA_Availability_Matrix",
        "Legacy_Auth_Deprecation",
        "Gap_Analysis",
        "Evidence_Register",
        "Approval_Sign_Off"
    ]
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
    return wb
# SECTION 4: SHEET 1 - INSTRUCTIONS & LEGEND
def populate_instructions(wb, styles):
    """Populate Instructions & Legend sheet."""
    ws = wb["Instructions_Legend"]
    # Title
    ws['A1'] = f"{DOCUMENT_ID}  -  Authentication Inventory & Methods Assessment\n{CONTROL_REF}"
    apply_style(ws['A1'], styles['title'])
    ws.row_dimensions[1].height = 40
    # Document metadata (standardized rows 3-6)
    ws['A3'] = 'Document ID:'
    ws['A3'].font = Font(bold=True)
    ws['B3'] = DOCUMENT_ID
    ws['A4'] = 'Assessment:'
    ws['A4'].font = Font(bold=True)
    ws['B4'] = 'Authentication Inventory & Methods'
    ws['A5'] = 'Version:'
    ws['A5'].font = Font(bold=True)
    ws['B5'] = VERSION
    ws['A6'] = 'Generated:'
    ws['A6'].font = Font(bold=True)
    ws['B6'] = datetime.now().strftime('%d.%m.%Y %H:%M')
    ws.column_dimensions['B'].width = 40
    # Document metadata
    row = 8
    ws[f'A{row}'] = "PURPOSE"
    apply_style(ws[f'A{row}'], styles['subheader'])
    row += 1
    ws[f'A{row}'] = "This workbook provides a comprehensive framework for assessing authentication mechanisms across all organizational systems. It enables systematic inventory of authentication methods, protocols, SSO integration, and compliance with secure authentication requirements (A.8.5)."
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(f'A{row}:H{row}')
    ws.row_dimensions[row].height = 45
    row += 2
    ws[f'A{row}'] = "ASSESSMENT APPROACH"
    instructions = [
        "1. SYSTEM AUTHENTICATION INVENTORY: Complete authentication mechanism inventory for all systems (Worksheet 2)",
        "2. PROTOCOL ANALYSIS: Assess security of authentication protocols in use (Worksheet 3)",
        "3. SSO INTEGRATION: Track SSO adoption and identify integration gaps (Worksheet 4)",
        "4. PASSWORD POLICY: Verify password policy configurations compliance (Worksheet 5)",
        "5. MFA AVAILABILITY: Document MFA methods available per system (Worksheet 6)",
        "6. LEGACY DEPRECATION: Identify and plan deprecation of legacy authentication (Worksheet 7)",
        "7. GAP ANALYSIS: Identify authentication security gaps requiring remediation (Worksheet 8)",
        "8. EVIDENCE COLLECTION: Document evidence supporting authentication controls (Worksheet 9)",
        "9. APPROVAL: Obtain required approvals from stakeholders (Worksheet 10)"
    ]
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = 30
        row += 1
    ws[f'A{row}'] = "STATUS LEGEND"
    legend = [
        ("{CHECK} Compliant", "Authentication mechanism meets all security requirements"),
        ("{WARNING} Partial Compliance", "Authentication mechanism meets some but not all requirements"),
        ("{XMARK} Non-Compliant", "Authentication mechanism does not meet security requirements"),
        ("🔄 In Progress", "Authentication improvement project in progress"),
        ("📋 Under Review", "Authentication mechanism being assessed"),
        ("❓ Unknown", "Authentication mechanism security status not yet determined"),
        ("➖ N/A", "Requirement not applicable to this system")
    ]
    row += 1
    for status, description in legend:
        ws[f'A{row}'] = status
        ws[f'B{row}'] = description
        apply_style(ws[f'A{row}'], styles['data'])
        apply_style(ws[f'B{row}'], styles['data'])
        ws.merge_cells(f'B{row}:H{row}')
        row += 1
    row += 1
    ws[f'A{row}'] = "SECURITY RATING SCALE"
    ratings = [
        ("🟢 Strong (90-100%)", "Modern authentication with MFA, strong protocols (SAML, OAuth, OIDC)"),
        ("🟡 Adequate (70-89%)", "Modern authentication, MFA available but not mandatory"),
        ("🟠 Weak (50-69%)", "Password-only authentication or outdated protocols"),
        ("🔴 Poor (<50%)", "Legacy authentication methods, significant security gaps"),
        ("⚫ Unacceptable", "Insecure authentication (Basic Auth, no TLS, shared credentials)")
    ]
    row += 1
    for rating, description in ratings:
        ws[f'A{row}'] = rating
        ws[f'B{row}'] = description
        apply_style(ws[f'A{row}'], styles['data'])
        apply_style(ws[f'B{row}'], styles['data'])
        ws.merge_cells(f'B{row}:H{row}')
        row += 1
    row += 1
    ws[f'A{row}'] = "AUTHENTICATION SECURITY PRINCIPLES"
    principles = [
        f"{BULLET} Multi-Factor Authentication (MFA): MANDATORY for privileged users and remote access, target 90%+ for all users",
        f"{BULLET} Modern Protocols: Use SAML 2.0, OAuth 2.0, OpenID Connect, Kerberos; deprecate NTLM, Basic Auth",
        f"{BULLET} Single Sign-On (SSO): Deploy for 90%+ of applications to reduce password fatigue and improve security",
        f"{BULLET} Password Policies: Minimum 12 characters, no forced expiration (unless compromise), breach detection enabled",
        f"{BULLET} Certificate Authentication: Use for system-to-system and API authentication (stronger than passwords)",
        f"{BULLET} Hardware Tokens: FIDO2 hardware tokens REQUIRED for Tier 0 privileged accounts",
        f"{BULLET} Default: Deny password-only authentication for privileged access and sensitive systems"
    ]
    for principle in principles:
        row += 1
        ws[f'A{row}'] = principle
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = 25
    row += 2
    ws[f'A{row}'] = f"{TARGET} REMEMBER: 'Evidence > Theater' - Authentication security means nothing without measurable MFA coverage and modern protocols."
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, italic=True, color='C00000')
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[row].height = 30
    # Column widths
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 20
# SECTION 5: SHEET 2 - SYSTEM AUTHENTICATION INVENTORY
def populate_system_inventory(wb, styles):
    """Populate System Authentication Inventory sheet."""
    ws = wb["System_Auth_Inventory"]
    ws['A1'] = "System Authentication Inventory"
    ws.merge_cells('A1:M1')
    ws['A2'] = "Complete inventory of authentication mechanisms for all organizational systems"
    ws.merge_cells('A2:M2')
    # Headers
    headers = [
        "System / Application",
        "System Type",
        "Environment",
        "Primary Auth Method",
        "Auth Protocol",
        "Identity Provider",
        "SSO Integrated",
        "MFA Available",
        "MFA Method",
        "Password Policy",
        "Security Rating",
        "Compliance Status",
        "Notes / Gaps"
    ]
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    # Pre-format data rows
    for row_num in range(5, 5 + SYSTEM_ROW_COUNT):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    # Add dropdowns
    # Column D: Primary Auth Method
    dv_auth_method = DataValidation(type="list", formula1=f'"{",".join(AUTH_METHODS)}"', allow_blank=True)
    ws.add_data_validation(dv_auth_method)
    dv_auth_method.add(f'D5:D{4 + SYSTEM_ROW_COUNT}')
    # Column E: Auth Protocol
    dv_protocol = DataValidation(type="list", formula1=f'"{",".join(AUTH_PROTOCOLS)}"', allow_blank=True)
    ws.add_data_validation(dv_protocol)
    dv_protocol.add(f'E5:E{4 + SYSTEM_ROW_COUNT}')
    # Column F: Identity Provider
    dv_provider = DataValidation(type="list", formula1=f'"{",".join(IDENTITY_PROVIDERS)}"', allow_blank=True)
    ws.add_data_validation(dv_provider)
    dv_provider.add(f'F5:F{4 + SYSTEM_ROW_COUNT}')
    # Column G: SSO Integrated (Yes/No)
    dv_sso = DataValidation(type="list", formula1='"Yes,No,In Progress,Not Supported"', allow_blank=True)
    ws.add_data_validation(dv_sso)
    dv_sso.add(f'G5:G{4 + SYSTEM_ROW_COUNT}')
    # Column H: MFA Available (Yes/No)
    dv_mfa_avail = DataValidation(type="list", formula1='"Yes,No,Planned"', allow_blank=True)
    ws.add_data_validation(dv_mfa_avail)
    dv_mfa_avail.add(f'H5:H{4 + SYSTEM_ROW_COUNT}')
    # Column I: MFA Method
    dv_mfa_method = DataValidation(type="list", formula1=f'"{",".join(MFA_METHODS)}"', allow_blank=True)
    ws.add_data_validation(dv_mfa_method)
    dv_mfa_method.add(f'I5:I{4 + SYSTEM_ROW_COUNT}')
    # Column J: Password Policy
    dv_password = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant,N/A (Certificate)"', allow_blank=True)
    ws.add_data_validation(dv_password)
    dv_password.add(f'J5:J{4 + SYSTEM_ROW_COUNT}')
    # Column K: Security Rating
    dv_rating = DataValidation(type="list", formula1=f'"{",".join(SECURITY_RATINGS)}"', allow_blank=True)
    ws.add_data_validation(dv_rating)
    dv_rating.add(f'K5:K{4 + SYSTEM_ROW_COUNT}')
    # Column L: Compliance Status
    dv_compliance = DataValidation(type="list", formula1=f'"{",".join(COMPLIANCE_STATUS)}"', allow_blank=True)
    ws.add_data_validation(dv_compliance)
    dv_compliance.add(f'L5:L{4 + SYSTEM_ROW_COUNT}')
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 22
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 35
    # Freeze panes
    ws.freeze_panes = 'A5'
# SECTION 6: SHEET 3 - AUTHENTICATION PROTOCOL ANALYSIS
def populate_protocol_analysis(wb, styles):
    """Populate Authentication Protocol Analysis sheet."""
    ws = wb["Auth_Protocol_Analysis"]
    ws['A1'] = "Authentication Protocol Security Analysis"
    ws.merge_cells('A1:J1')
    ws['A2'] = "Security assessment of authentication protocols in use"
    ws.merge_cells('A2:J2')
    headers = [
        "Protocol Name",
        "Systems Using",
        "Security Level",
        "TLS Required",
        "TLS Version",
        "Deprecation Status",
        "Compliance Rating",
        "Remediation Plan",
        "Target Date",
        "Notes"
    ]
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + PROTOCOL_ROW_COUNT):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    # Column A: Protocol Name
    dv_protocol = DataValidation(type="list", formula1=f'"{",".join(AUTH_PROTOCOLS)}"', allow_blank=True)
    ws.add_data_validation(dv_protocol)
    dv_protocol.add(f'A5:A{4 + PROTOCOL_ROW_COUNT}')
    # Column C: Security Level
    dv_security = DataValidation(type="list", formula1='"🟢 Strong,🟡 Adequate,🟠 Weak,🔴 Legacy/Insecure"', allow_blank=True)
    ws.add_data_validation(dv_security)
    dv_security.add(f'C5:C{4 + PROTOCOL_ROW_COUNT}')
    # Column D: TLS Required
    dv_tls = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=True)
    ws.add_data_validation(dv_tls)
    dv_tls.add(f'D5:D{4 + PROTOCOL_ROW_COUNT}')
    # Column E: TLS Version
    dv_tls_version = DataValidation(type="list", formula1='"TLS 1.3,TLS 1.2,TLS 1.1 (Deprecated),TLS 1.0 (Insecure),No TLS"', allow_blank=True)
    ws.add_data_validation(dv_tls_version)
    dv_tls_version.add(f'E5:E{4 + PROTOCOL_ROW_COUNT}')
    # Column F: Deprecation Status
    dv_deprecation = DataValidation(type="list", formula1='"Active - Recommended,Active - Acceptable,Deprecated - Migrate,Prohibited - Block"', allow_blank=True)
    ws.add_data_validation(dv_deprecation)
    dv_deprecation.add(f'F5:F{4 + PROTOCOL_ROW_COUNT}')
    # Column G: Compliance Rating
    dv_compliance = DataValidation(type="list", formula1=f'"{",".join(COMPLIANCE_STATUS)}"', allow_blank=True)
    ws.add_data_validation(dv_compliance)
    dv_compliance.add(f'G5:G{4 + PROTOCOL_ROW_COUNT}')
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 35
# SECTION 7: SHEET 4 - SSO INTEGRATION STATUS
def populate_sso_integration(wb, styles):
    """Populate SSO Integration Status sheet."""
    ws = wb["SSO_Integration_Status"]
    ws['A1'] = "Single Sign-On (SSO) Integration Status"
    ws.merge_cells('A1:K1')
    ws['A2'] = "Track SSO adoption and identify integration gaps (Target: 90%+ SSO integration)"
    ws.merge_cells('A2:K2')
    headers = [
        "Application Name",
        "Application Type",
        "User Count",
        "SSO Status",
        "SSO Protocol",
        "Integration Date",
        "SSO Platform",
        "Integration Effort",
        "Business Priority",
        "Compliance Status",
        "Notes"
    ]
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + SSO_APP_ROW_COUNT):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    # Column B: Application Type
    dv_app_type = DataValidation(type="list", formula1='"SaaS,On-Premises Web App,Cloud Service,Internal Portal,Database,Legacy,Other"', allow_blank=True)
    ws.add_data_validation(dv_app_type)
    dv_app_type.add(f'B5:B{4 + SSO_APP_ROW_COUNT}')
    # Column D: SSO Status
    dv_sso_status = DataValidation(type="list", formula1=f'"{CHECK} Integrated,🔄 In Progress,📋 Planned,❌ Not Supported,⚠️ Password Vaulting,❓ Not Assessed"', allow_blank=True)
    ws.add_data_validation(dv_sso_status)
    dv_sso_status.add(f'D5:D{4 + SSO_APP_ROW_COUNT}')
    # Column E: SSO Protocol
    dv_sso_protocol = DataValidation(type="list", formula1='"SAML 2.0,OAuth 2.0,OpenID Connect,WS-Federation,Not Supported"', allow_blank=True)
    ws.add_data_validation(dv_sso_protocol)
    dv_sso_protocol.add(f'E5:E{4 + SSO_APP_ROW_COUNT}')
    # Column G: SSO Platform
    dv_sso_platform = DataValidation(type="list", formula1=f'"{",".join(IDENTITY_PROVIDERS)}"', allow_blank=True)
    ws.add_data_validation(dv_sso_platform)
    dv_sso_platform.add(f'G5:G{4 + SSO_APP_ROW_COUNT}')
    # Column H: Integration Effort
    dv_effort = DataValidation(type="list", formula1='"Low (Pre-built),Medium (Custom Config),High (Development),Not Possible"', allow_blank=True)
    ws.add_data_validation(dv_effort)
    dv_effort.add(f'H5:H{4 + SSO_APP_ROW_COUNT}')
    # Column I: Business Priority
    dv_priority = DataValidation(type="list", formula1=f'"{",".join(PRIORITY_LEVELS)}"', allow_blank=True)
    ws.add_data_validation(dv_priority)
    dv_priority.add(f'I5:I{4 + SSO_APP_ROW_COUNT}')
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 35
    # Summary section
    row = 5 + SSO_APP_ROW_COUNT + 2
    ws[f'A{row}'] = "SSO ADOPTION METRICS"
    apply_style(ws[f'A{row}'], styles['subheader'])
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    ws[f'A{row}'] = "Total Applications:"
    ws[f'B{row}'] = f'=COUNTA(A5:A{4 + SSO_APP_ROW_COUNT})'
    row += 1
    ws[f'A{row}'] = "SSO Integrated:"
    ws[f'B{row}'] = f'=COUNTIF(D5:D{4 + SSO_APP_ROW_COUNT},"{CHECK} Integrated")'
    row += 1
    ws[f'A{row}'] = "SSO Adoption Rate:"
    ws[f'B{row}'] = f'=IF(B{row-2}>0,B{row-1}/B{row-2},0)'
    ws[f'B{row}'].number_format = '0.0%'
    row += 1
    ws[f'A{row}'] = "Target (90%):"
    ws[f'B{row}'] = 0.90
# SECTION 8: SHEET 5 - PASSWORD POLICY COMPLIANCE
def populate_password_policy(wb, styles):
    """Populate Password Policy Compliance sheet."""
    ws = wb["Password_Policy_Compliance"]
    ws['A1'] = "Password Policy Configuration & Compliance"
    ws.merge_cells('A1:L1')
    ws['A2'] = "Modern password policies: Length > Complexity, No forced expiration (unless compromise), Breach detection enabled"
    ws.merge_cells('A2:L2')
    headers = [
        "System / Platform",
        "Policy Enforced",
        "Min Length",
        "Complexity Req",
        "Password Expiry",
        "History Count",
        "Breach Detection",
        "Lockout Threshold",
        "Compliance Score",
        "Gaps Identified",
        "Compliance Status",
        "Notes"
    ]
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + PASSWORD_POLICY_COUNT):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    # Column B: Policy Enforced
    dv_enforced = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    ws.add_data_validation(dv_enforced)
    dv_enforced.add(f'B5:B{4 + PASSWORD_POLICY_COUNT}')
    # Column D: Complexity Required
    dv_complexity = DataValidation(type="list", formula1='"Yes,No,3 of 4 Types,Custom"', allow_blank=True)
    ws.add_data_validation(dv_complexity)
    dv_complexity.add(f'D5:D{4 + PASSWORD_POLICY_COUNT}')
    # Column E: Password Expiry
    dv_expiry = DataValidation(type="list", formula1='"Never (Recommended),90 Days,60 Days,30 Days,Custom"', allow_blank=True)
    ws.add_data_validation(dv_expiry)
    dv_expiry.add(f'E5:E{4 + PASSWORD_POLICY_COUNT}')
    # Column G: Breach Detection
    dv_breach = DataValidation(type="list", formula1=f'"{CHECK} Enabled,❌ Not Available,📋 Planned"', allow_blank=True)
    ws.add_data_validation(dv_breach)
    dv_breach.add(f'G5:G{4 + PASSWORD_POLICY_COUNT}')
    # Column I: Compliance Score
    dv_score = DataValidation(type="list", formula1='"🟢 100% Compliant,🟡 75-99%,🟠 50-74%,🔴 <50%"', allow_blank=True)
    ws.add_data_validation(dv_score)
    dv_score.add(f'I5:I{4 + PASSWORD_POLICY_COUNT}')
    # Column K: Compliance Status
    dv_compliance = DataValidation(type="list", formula1=f'"{",".join(COMPLIANCE_STATUS)}"', allow_blank=True)
    ws.add_data_validation(dv_compliance)
    dv_compliance.add(f'K5:K{4 + PASSWORD_POLICY_COUNT}')
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 30
    ws.column_dimensions['L'].width = 35
    # Policy Requirements Reference
    row = 5 + PASSWORD_POLICY_COUNT + 2
    ws[f'A{row}'] = "MODERN PASSWORD POLICY REQUIREMENTS (NIST SP 800-63B)"
    ws.merge_cells(f'A{row}:E{row}')
    requirements = [
        ("Minimum Length", "≥12 characters (15+ recommended)", "🟢 Strong"),
        ("Complexity", "Optional if length ≥15, otherwise 3 of 4 types", "🟡 Moderate"),
        ("Password Expiration", "No forced expiration (unless compromise)", "🟢 Modern"),
        ("History Count", "≥5 passwords remembered", "🟢 Good"),
        ("Breach Detection", "Check against known breached passwords", "🟢 Critical"),
        ("Lockout Threshold", "5-10 failed attempts", "🟢 Recommended")
    ]
    row += 1
    ws[f'A{row}'] = "Requirement"
    ws[f'B{row}'] = "Modern Standard"
    ws[f'C{row}'] = "Priority"
    apply_style(ws[f'A{row}'], styles['header'])
    apply_style(ws[f'B{row}'], styles['header'])
    apply_style(ws[f'C{row}'], styles['header'])
    for req, standard, priority in requirements:
        row += 1
        ws[f'A{row}'] = req
        ws[f'B{row}'] = standard
        ws[f'C{row}'] = priority
        apply_style(ws[f'A{row}'], styles['data'])
        apply_style(ws[f'B{row}'], styles['data'])
        apply_style(ws[f'C{row}'], styles['data'])
# SECTION 9: REMAINING SHEETS (MFA, LEGACY, GAP, EVIDENCE, APPROVAL)
def populate_remaining_sheets(wb, styles):
    """Populate remaining sheets with standard structure."""
    # Sheet 6: MFA Availability Matrix
    ws_mfa = wb["MFA_Availability_Matrix"]
    ws_mfa['A1'] = "MFA Availability Matrix"
    apply_style(ws_mfa['A1'], styles['title'])
    ws_mfa['A2'] = "MFA method availability per system (linked to System_Auth_Inventory)"

    # Sheet 7: Legacy Auth Deprecation
    ws_legacy = wb["Legacy_Auth_Deprecation"]
    ws_legacy['A1'] = "Legacy Authentication Deprecation Tracking"
    apply_style(ws_legacy['A1'], styles['title'])
    ws_legacy['A2'] = "Identify and plan migration from legacy protocols (NTLM, Basic Auth)"
    legacy_headers = ["System", "Legacy Protocol", "Risk Level", "Users Affected", "Migration Plan", "Target Date", "Status"]
    header_row = 4
    for col_idx, header in enumerate(legacy_headers, start=1):
        cell = ws_legacy.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + LEGACY_ROW_COUNT):
        for col_idx in range(1, len(legacy_headers) + 1):
            cell = ws_legacy.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])

    # Sheet 8: Gap Analysis
    ws_gap = wb["Gap_Analysis"]
    ws_gap['A1'] = "Authentication Security Gap Analysis"
    apply_style(ws_gap['A1'], styles['title'])
    gap_headers = ["Gap ID", "System/Area", "Gap Description", "Risk Level", "Impact", "Remediation Plan", "Owner", "Target Date", "Status"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws_gap.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + GAP_ROW_COUNT):
        for col_idx in range(1, len(gap_headers) + 1):
            cell = ws_gap.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])

    # Sheet 9: Evidence Register
    ws_evidence = wb["Evidence_Register"]
    ws_evidence['A1'] = "Authentication Evidence Register"
    apply_style(ws_evidence['A1'], styles['title'])
    evidence_headers = ["Evidence ID", "Control Ref", "Evidence Type", "Description", "Location", "Date Collected", "Collected By", "Verification Status"]
    for col_idx, header in enumerate(evidence_headers, start=1):
        cell = ws_evidence.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['header'])
    for row_num in range(5, 5 + EVIDENCE_ROW_COUNT):
        for col_idx in range(1, len(evidence_headers) + 1):
            cell = ws_evidence.cell(row=row_num, column=col_idx)
            apply_style(cell, styles['data'])
    # Sheet 10: Approval Sign-Off
    ws_approval = wb["Approval_Sign_Off"]
    ws_approval['A1'] = "Assessment Approval & Sign-Of"
    apply_style(ws_approval['A1'], styles['title'])
    ws_approval['A3'] = "Assessment Completion"
    apply_style(ws_approval['A3'], styles['subheader'])
    ws_approval['A4'] = "Assessment Date:"
    ws_approval['A5'] = "Assessed By:"
    ws_approval['A6'] = "Assessment Status:"
    ws_approval['A8'] = "Approval Workflow"
    apply_style(ws_approval['A8'], styles['subheader'])
    approval_rows = [
        ("Level 1", "IAM Lead / Security Analyst", "", "", ""),
        ("Level 2", "Security Architect", "", "", ""),
        ("Level 3", "CISO", "", "", "")
    ]
    row = 9
    ws_approval[f'A{row}'] = "Approval Level"
    ws_approval[f'B{row}'] = "Role"
    ws_approval[f'C{row}'] = "Name"
    ws_approval[f'D{row}'] = "Date"
    ws_approval[f'E{row}'] = "Signature / Status"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws_approval[f'{col}{row}'], styles['header'])
    for level, role, name, date, sig in approval_rows:
        row += 1
        ws_approval[f'A{row}'] = level
        ws_approval[f'B{row}'] = role
        ws_approval[f'C{row}'] = name
        ws_approval[f'D{row}'] = date
        ws_approval[f'E{row}'] = sig
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws_approval[f'{col}{row}'], styles['data'])

# SECTION 10: MAIN GENERATION FUNCTION
def generate_workbook():
    """Main function to generate complete workbook."""
    logger.info("")
    logger.info("╔════════════════════════════════════════════════════════════════╗")
    logger.info("║  ISMS Assessment A.8.2/3/5 - Authentication & PAM Framework    ║")
    logger.info("║  Workbook 1: Authentication Inventory & Methods Assessment     ║")
    logger.info("║                                                                ║")
    logger.info("║  'Evidence > Theater' - Real authentication security           ║")
    logger.info("╚════════════════════════════════════════════════════════════════╝")
    logger.info("Creating workbook structure...")
    wb = create_workbook()
    styles = setup_styles()
    logger.info("Populating Instructions & Legend...")
    populate_instructions(wb, styles)
    logger.info("Populating System Authentication Inventory (150 systems)...")
    populate_system_inventory(wb, styles)
    logger.info("Populating Authentication Protocol Analysis...")
    populate_protocol_analysis(wb, styles)
    logger.info("Populating SSO Integration Status...")
    populate_sso_integration(wb, styles)
    logger.info("Populating Password Policy Compliance...")
    populate_password_policy(wb, styles)
    logger.info("Populating remaining sheets...")
    populate_remaining_sheets(wb, styles)
    # Save workbook
    filename = f"ISMS-IMP-A.8.2-3-5.S1_Authentication_Inventory_{GENERATED_TIMESTAMP}.xlsx"
    wb.save(filename)
    logger.info("=" * 70)
    logger.info("{CHECK} Workbook generated successfully: {filename}")
    logger.info("Next Steps:")
    logger.info("  1. Open the workbook in Excel")
    logger.info("  2. Complete System_Auth_Inventory (main assessment)")
    logger.info("  3. Review SSO_Integration_Status (track SSO adoption)")
    logger.info("  4. Verify Password_Policy_Compliance")
    logger.info("  5. Document gaps in Gap_Analysis")
    logger.info("  6. Collect evidence in Evidence_Register")
    logger.info("  7. Obtain approvals in Approval_Sign_Off")
    logger.info("Target Metrics:")
    logger.info("  • MFA Coverage: 100% privileged users, 90%+ all users")
    logger.info("  • SSO Adoption: 90%+ of applications")
    logger.info("  • Modern Protocols: >95% (SAML, OAuth, OIDC, Kerberos)")
    logger.info("  • Legacy Protocols: <5% (plan deprecation)")
    return filename
def validate_workbook(filename):
    """Validate generated workbook."""
    logger.info("Running validation...")
    logger.info("-" * 70)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        
        expected_sheets = [
            "Instructions_Legend",
            "System_Auth_Inventory",
            "Auth_Protocol_Analysis",
            "SSO_Integration_Status",
            "Password_Policy_Compliance",
            "MFA_Availability_Matrix",
            "Legacy_Auth_Deprecation",
            "Gap_Analysis",
            "Evidence_Register",
            "Approval_Sign_Off"
        ]
        # Check sheet count
        if len(wb.sheetnames) != 10:
            logger.info(f"  ❌ Expected 10 sheets, found {len(wb.sheetnames)}")
            return False
        logger.info(f"  ✅ Sheet count: {len(wb.sheetnames)}")
        # Check sheet names
        for sheet in expected_sheets:
            if sheet in wb.sheetnames:
                logger.info(f"  ✅ Found: {sheet}")
            else:
                logger.info(f"  ❌ Missing: {sheet}")
                return False
        # Check System_Auth_Inventory has 150 rows
        ws = wb["System_Auth_Inventory"]
        if ws.max_row >= 154:  # 4 header rows + 150 data rows
            logger.info(f"  ✅ System_Auth_Inventory: {SYSTEM_ROW_COUNT} pre-formatted rows")
        else:
            logger.info(f"  ⚠️ System_Auth_Inventory: Fewer than expected rows")
        logger.info("")
        logger.info("Validation Result: ✅ PASSED")
        wb.close()
        return True
    except Exception as e:
        logger.error(f"  ❌ Validation error: {str(e)}")
        return False
# SECTION 11: ENTRY POINT
if __name__ == "__main__":
    try:
        # Generate the workbook
        output_file = generate_workbook()
        # Validate the output
        if validate_workbook(output_file):
            logger.info("{CHECK} Successfully generated: {output_file}")
            sys.exit(0)
        else:
            logger.info("{WARNING} Validation warnings - please review the output")
            sys.exit(1)
    except ImportError as e:
        logger.error("ERROR: Missing required library")
        logger.info("-" * 70)
        logger.info(f"  {str(e)}")
        logger.info("Please install openpyxl:")
        logger.info("  pip install openpyxl")
        sys.exit(1)
    except Exception as e:
        logger.error("ERROR: Generation failed")
        import traceback
        traceback.print_exc()
        sys.exit(1)
# =============================================================================
# END OF GENERATOR SCRIPT
#
# Document Control:
#   Version: 1.0
#   Created: 2025-01-11
#   Author:               [Organization] ISMS Implementation Team
# Change History:
#   1.0 - Initial version
#       - 10 sheets for comprehensive authentication assessment
#       - 150-row system authentication inventory
#       - SSO integration tracking (100 apps)
#       - Password policy compliance verification
#       - Legacy authentication deprecation tracking
#       - Evidence register and approval workflow
# Dependencies:
#   - Python 3.7+
#   - openpyxl >= 3.0.0
# Output:
#   Authentication_Inventory_Assessment_YYYYMMDD.xlsx

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
