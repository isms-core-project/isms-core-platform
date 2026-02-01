#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.33-34.2 - Audit Activity Management Assessment
================================================================================

ISO/IEC 27001:2022 Controls A.8.33: Test Information & A.8.34: Protection of
Information Systems During Audit Testing
Assessment Domain 2 of 3: Audit Activity Governance & System Protection

This script generates a comprehensive Excel assessment workbook for managing
audit testing governance including activity planning, tool authorization,
access control, disruption mitigation, and evidence protection.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.33-34.2"
WORKBOOK_NAME = "Audit Activity Management Assessment"
CONTROL_ID = "A.8.33-34"
CONTROL_NAME = "Testing and Audit Protection"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions and Legend sheet."""
    ws.title = "Instructions_and_Legend"

    instructions = [
        ["ISMS-IMP-A.8.33-34.2 - Audit Activity Management Assessment"],
        ["ISO/IEC 27001:2022 - Control A.8.34: Protection of Information Systems During Audit Testing"],
        [""],
        ["PURPOSE"],
        ["This workbook tracks audit testing governance and system protection compliance including"],
        ["audit activity register, tool authorization, access tracking, disruption mitigation, and evidence protection."],
        [""],
        ["SHEETS"],
        ["1. Instructions_and_Legend - This guidance sheet"],
        ["2. Audit_Activity_Register - Registry of all planned and completed audit activities"],
        ["3. Audit_Tool_Authorization - Governance of audit tools and testing techniques"],
        ["4. Audit_Access_Tracking - Management of auditor access to systems and data"],
        ["5. Disruption_Mitigation_Plans - Protection measures for production systems during testing"],
        ["6. Audit_Evidence_Protection - Secure handling and retention of audit evidence"],
        ["7. Summary_Dashboard - Executive overview of audit governance compliance"],
        ["8. Evidence_Register - Documentation supporting assessment findings"],
        ["9. Approval_Sign_Off - Three-level approval workflow"],
        [""],
        ["STATUS LEGEND"],
        ["Compliant - Fully meets requirement"],
        ["Partial - Partially meets requirement, improvement needed"],
        ["Non-Compliant - Does not meet requirement"],
        ["Planned - Implementation planned"],
        ["N/A - Not Applicable"],
        [""],
        ["RISK LEVEL SCALE"],
        ["High - Could cause significant disruption or data exposure (CISO + Business Owner approval)"],
        ["Medium - Moderate risk to operations (IT Security Manager approval)"],
        ["Low - Minimal risk, standard tools (Security Team approval)"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14, color="003366")
            elif row_num == 2:
                cell.font = Font(bold=True, size=11, color="003366")
            elif value in ["PURPOSE", "SHEETS", "STATUS LEGEND", "RISK LEVEL SCALE"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 90


def create_audit_activity_register_sheet(ws):
    """Create the Audit Activity Register sheet."""
    ws.title = "Audit_Activity_Register"

    ws.merge_cells('A1:V1')
    title_cell = ws.cell(row=1, column=1, value="AUDIT ACTIVITY REGISTER")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Registry of all planned and completed audit activities")

    headers = [
        "Audit_ID", "Audit_Name", "Audit_Type", "Audit_Scope", "Audit_Firm_Team",
        "Lead_Auditor", "Planned_Start", "Planned_End", "Actual_Start", "Actual_End",
        "Audit_Status", "Management_Approval", "Approver", "Approval_Date",
        "Systems_in_Scope", "Data_Access_Required", "Testing_Type", "Findings_Count",
        "Critical_Findings", "Report_Location", "Follow_up_Status", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(type="list", formula1='"Internal,External-Financial,External-SOC2,Penetration Test,Vulnerability Assessment,Red Team,Regulatory,Compliance"')
    ws.add_data_validation(type_dv)
    type_dv.add('C5:C54')

    status_dv = DataValidation(type="list", formula1='"Planned,In Progress,Completed,Cancelled,On Hold"')
    ws.add_data_validation(status_dv)
    status_dv.add('K5:K54')

    approval_dv = DataValidation(type="list", formula1='"Approved,Pending,Not Required"')
    ws.add_data_validation(approval_dv)
    approval_dv.add('L5:L54')

    testing_dv = DataValidation(type="list", formula1='"Non-Invasive,Read-Only,Active Testing,Exploitation"')
    ws.add_data_validation(testing_dv)
    testing_dv.add('Q5:Q54')

    followup_dv = DataValidation(type="list", formula1='"Open,In Progress,Closed"')
    ws.add_data_validation(followup_dv)
    followup_dv.add('U5:U54')

    # Format input rows
    for row in range(5, 55):
        ws.cell(row=row, column=1, value=f"AUD-{str(row-4).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics
    ws.cell(row=57, column=1, value="AUDIT STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Audits", "=COUNTA(B5:B54)"),
        ("Planned Audits", '=COUNTIF(K5:K54,"Planned")'),
        ("In Progress", '=COUNTIF(K5:K54,"In Progress")'),
        ("Completed", '=COUNTIF(K5:K54,"Completed")'),
        ("Approved Audits", '=COUNTIF(L5:L54,"Approved")'),
        ("Pending Approval", '=COUNTIF(L5:L54,"Pending")'),
        ("Internal Audits", '=COUNTIF(C5:C54,"Internal")'),
        ("Penetration Tests", '=COUNTIF(C5:C54,"Penetration Test")'),
        ("Total Findings", "=SUM(R5:R54)"),
        ("Total Critical Findings", "=SUM(S5:S54)"),
        ("Open Follow-ups", '=COUNTIF(U5:U54,"Open")'),
    ]

    row = 59
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [12, 35, 25, 50, 30, 25, 12, 12, 12, 12, 18, 18, 25, 12, 40, 30, 25, 12, 12, 35, 18, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_audit_tool_authorization_sheet(ws):
    """Create the Audit Tool Authorization sheet."""
    ws.title = "Audit_Tool_Authorization"

    ws.merge_cells('A1:U1')
    title_cell = ws.cell(row=1, column=1, value="AUDIT TOOL AUTHORIZATION")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Registry of authorized audit and testing tools")

    headers = [
        "Tool_ID", "Tool_Name", "Tool_Version", "Tool_Category", "Vendor_Source",
        "Tool_Owner", "Authorization_Status", "Authorization_Date", "Authorized_By",
        "Risk_Level", "Authorized_Use_Cases", "Restrictions", "Required_Approvals",
        "Storage_Location", "Access_Restricted_To", "Last_Security_Review",
        "Next_Review_Due", "Usage_Logging_Required", "License_Status",
        "License_Expiry", "Evidence_Reference"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    category_dv = DataValidation(type="list", formula1='"Vulnerability Scanner,Penetration Tool,Network Analyzer,Web App Scanner,Forensic Tool,Credential Tester,Exploitation Framework,Other"')
    ws.add_data_validation(category_dv)
    category_dv.add('D5:D54')

    auth_dv = DataValidation(type="list", formula1='"Authorized,Pending,Unauthorized,Prohibited"')
    ws.add_data_validation(auth_dv)
    auth_dv.add('G5:G54')

    risk_dv = DataValidation(type="list", formula1='"High,Medium,Low"')
    ws.add_data_validation(risk_dv)
    risk_dv.add('J5:J54')

    yn_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('R5:R54')

    license_dv = DataValidation(type="list", formula1='"Valid,Expired,N/A"')
    ws.add_data_validation(license_dv)
    license_dv.add('S5:S54')

    # Format input rows
    for row in range(5, 55):
        ws.cell(row=row, column=1, value=f"TOOL-{str(row-4).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics
    ws.cell(row=57, column=1, value="TOOL RISK SUMMARY").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Tools", "=COUNTA(B5:B54)"),
        ("Authorized Tools", '=COUNTIF(G5:G54,"Authorized")'),
        ("Pending Authorization", '=COUNTIF(G5:G54,"Pending")'),
        ("Unauthorized Tools", '=COUNTIF(G5:G54,"Unauthorized")'),
        ("Prohibited Tools", '=COUNTIF(G5:G54,"Prohibited")'),
        ("High Risk Tools", '=COUNTIF(J5:J54,"High")'),
        ("Medium Risk Tools", '=COUNTIF(J5:J54,"Medium")'),
        ("Low Risk Tools", '=COUNTIF(J5:J54,"Low")'),
        ("Expired Licenses", '=COUNTIF(S5:S54,"Expired")'),
    ]

    row = 59
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [12, 30, 15, 25, 25, 25, 18, 12, 25, 15, 50, 40, 30, 35, 30, 12, 12, 12, 18, 12, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_audit_access_tracking_sheet(ws):
    """Create the Audit Access Tracking sheet."""
    ws.title = "Audit_Access_Tracking"

    ws.merge_cells('A1:W1')
    title_cell = ws.cell(row=1, column=1, value="AUDIT ACCESS TRACKING")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Registry of auditor access to systems and data")

    headers = [
        "Access_ID", "Auditor_Name", "Auditor_Organization", "Associated_Audit",
        "Access_Type", "Systems_Accessed", "Data_Classification_Accessed",
        "Access_Requested_Date", "Access_Start_Date", "Access_End_Date",
        "Actual_Revocation_Date", "Access_Status", "Approval_Status", "Approver",
        "Approval_Date", "Access_Logging_Enabled", "NDA_Signed", "NDA_Reference",
        "Supervision_Required", "Supervisor", "Multi_Factor_Auth_Required",
        "Revocation_Confirmation", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    access_type_dv = DataValidation(type="list", formula1='"Read-Only,Read-Write,Admin,Physical,Remote-VPN"')
    ws.add_data_validation(access_type_dv)
    access_type_dv.add('E5:E104')

    classification_dv = DataValidation(type="list", formula1='"Public,Internal,Confidential,Restricted"')
    ws.add_data_validation(classification_dv)
    classification_dv.add('G5:G104')

    status_dv = DataValidation(type="list", formula1='"Active,Revoked,Expired"')
    ws.add_data_validation(status_dv)
    status_dv.add('L5:L104')

    approval_dv = DataValidation(type="list", formula1='"Approved,Pending,Denied"')
    ws.add_data_validation(approval_dv)
    approval_dv.add('M5:M104')

    ynp_dv = DataValidation(type="list", formula1='"Yes,Partial,No"')
    ws.add_data_validation(ynp_dv)
    ynp_dv.add('P5:P104')

    yn_dv = DataValidation(type="list", formula1='"Yes,No,N/A"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('Q5:Q104')
    yn_dv.add('S5:S104')
    yn_dv.add('U5:U104')

    # Format input rows
    for row in range(5, 105):
        ws.cell(row=row, column=1, value=f"ACC-{str(row-4).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics
    ws.cell(row=107, column=1, value="ACCESS STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Access Grants", "=COUNTA(B5:B104)"),
        ("Active Access", '=COUNTIF(L5:L104,"Active")'),
        ("Revoked Access", '=COUNTIF(L5:L104,"Revoked")'),
        ("Expired Access", '=COUNTIF(L5:L104,"Expired")'),
        ("Approved Access", '=COUNTIF(M5:M104,"Approved")'),
        ("Pending Approval", '=COUNTIF(M5:M104,"Pending")'),
        ("Read-Only Access", '=COUNTIF(E5:E104,"Read-Only")'),
        ("Admin Access", '=COUNTIF(E5:E104,"Admin")'),
        ("Restricted Data Access", '=COUNTIF(G5:G104,"Restricted")'),
        ("NDA Coverage (Yes)", '=COUNTIF(Q5:Q104,"Yes")'),
        ("Access Logging (Yes)", '=COUNTIF(P5:P104,"Yes")'),
        ("MFA Required (Yes)", '=COUNTIF(U5:U104,"Yes")'),
    ]

    row = 109
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [12, 25, 25, 20, 20, 40, 18, 12, 12, 12, 12, 18, 18, 25, 12, 12, 12, 25, 12, 25, 12, 20, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_disruption_mitigation_sheet(ws):
    """Create the Disruption Mitigation Plans sheet."""
    ws.title = "Disruption_Mitigation_Plans"

    ws.merge_cells('A1:W1')
    title_cell = ws.cell(row=1, column=1, value="DISRUPTION MITIGATION PLANS")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Protection measures for production systems during audit testing")

    headers = [
        "System_ID", "System_Name", "System_Criticality", "Business_Owner",
        "Technical_Owner", "Associated_Audits", "Primary_Mitigation_Strategy",
        "Testing_Restrictions", "Permitted_Testing_Window", "Maximum_Test_Duration",
        "Backup_Required_Before_Test", "Recovery_Point_Objective", "Recovery_Time_Objective",
        "Rollback_Procedure_Location", "Rollback_Last_Tested", "Escalation_Contact",
        "Escalation_Phone", "Secondary_Contact", "Monitoring_Enhancement",
        "Incident_Response_Plan", "Last_Mitigation_Review", "Review_Due_Date", "Evidence_Reference"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    criticality_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(criticality_dv)
    criticality_dv.add('C5:C54')

    strategy_dv = DataValidation(type="list", formula1='"Staging First,Off-Hours,Rate Limiting,Scope Exclusion,Enhanced Monitoring,Standby Recovery,Read-Only Only,Multi-Strategy"')
    ws.add_data_validation(strategy_dv)
    strategy_dv.add('G5:G54')

    backup_dv = DataValidation(type="list", formula1='"Yes,No,Already Scheduled"')
    ws.add_data_validation(backup_dv)
    backup_dv.add('K5:K54')

    # Format input rows
    for row in range(5, 55):
        ws.cell(row=row, column=1, value=f"SYS-{str(row-4).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics
    ws.cell(row=57, column=1, value="MITIGATION STRATEGY SUMMARY").font = Font(bold=True, size=12)

    strategies = ["Staging First", "Off-Hours", "Rate Limiting", "Scope Exclusion", "Enhanced Monitoring", "Standby Recovery", "Read-Only Only", "Multi-Strategy"]
    row = 59
    for strategy in strategies:
        ws.cell(row=row, column=1, value=strategy).border = THIN_BORDER
        ws.cell(row=row, column=2, value=f'=COUNTIF(G5:G54,"{strategy}")').border = THIN_BORDER
        row += 1

    ws.cell(row=row + 1, column=1, value="SYSTEM COVERAGE STATISTICS").font = Font(bold=True, size=12)
    row += 3

    coverage_items = [
        ("Total Systems", "=COUNTA(B5:B54)"),
        ("Critical Systems", '=COUNTIF(C5:C54,"Critical")'),
        ("High Criticality", '=COUNTIF(C5:C54,"High")'),
        ("Medium Criticality", '=COUNTIF(C5:C54,"Medium")'),
        ("Low Criticality", '=COUNTIF(C5:C54,"Low")'),
        ("With Rollback Procedure", '=COUNTA(N5:N54)'),
        ("Backup Required", '=COUNTIF(K5:K54,"Yes")'),
    ]

    for label, formula in coverage_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [12, 30, 15, 25, 25, 30, 35, 50, 25, 15, 12, 15, 15, 35, 12, 25, 20, 25, 30, 35, 12, 12, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_audit_evidence_protection_sheet(ws):
    """Create the Audit Evidence Protection sheet."""
    ws.title = "Audit_Evidence_Protection"

    ws.merge_cells('A1:W1')
    title_cell = ws.cell(row=1, column=1, value="AUDIT EVIDENCE PROTECTION")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Secure handling and retention of audit evidence")

    headers = [
        "Evidence_Category_ID", "Evidence_Category", "Evidence_Description",
        "Sensitivity_Classification", "Example_Documents", "Primary_Storage_Location",
        "Backup_Location", "Encryption_at_Rest", "Encryption_in_Transit",
        "Access_Control_Type", "Authorized_Accessors", "Retention_Period",
        "Retention_Start_Event", "Destruction_Method", "Destruction_Approval",
        "Chain_of_Custody_Required", "Chain_of_Custody_Process", "Legal_Hold_Applicable",
        "Legal_Hold_Contact", "Integrity_Verification", "Last_Access_Review",
        "Evidence_Owner", "Evidence_Reference"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    classification_dv = DataValidation(type="list", formula1='"Public,Internal,Confidential,Restricted"')
    ws.add_data_validation(classification_dv)
    classification_dv.add('D5:D34')

    enc_rest_dv = DataValidation(type="list", formula1='"AES-256,AES-128,BitLocker,None,Other"')
    ws.add_data_validation(enc_rest_dv)
    enc_rest_dv.add('H5:H34')

    enc_transit_dv = DataValidation(type="list", formula1='"TLS 1.3,TLS 1.2,VPN,None,Other"')
    ws.add_data_validation(enc_transit_dv)
    enc_transit_dv.add('I5:I34')

    access_dv = DataValidation(type="list", formula1='"RBAC,ACL,Individual Permissions,None"')
    ws.add_data_validation(access_dv)
    access_dv.add('J5:J34')

    retention_dv = DataValidation(type="list", formula1='"1 Year,2 Years,3 Years,5 Years,7 Years,Permanent,Legal Hold"')
    ws.add_data_validation(retention_dv)
    retention_dv.add('L5:L34')

    destruction_dv = DataValidation(type="list", formula1='"Secure Delete,Shredding,Degaussing,Certified Destruction"')
    ws.add_data_validation(destruction_dv)
    destruction_dv.add('N5:N34')

    yn_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('P5:P34')

    legal_dv = DataValidation(type="list", formula1='"Yes,No,Potentially"')
    ws.add_data_validation(legal_dv)
    legal_dv.add('R5:R34')

    integrity_dv = DataValidation(type="list", formula1='"Hashing,Digital Signatures,Checksums,None"')
    ws.add_data_validation(integrity_dv)
    integrity_dv.add('T5:T34')

    # Pre-populate standard evidence categories
    categories = [
        ("EVCAT-001", "Penetration Test Reports", "Technical findings from pen tests", "Restricted"),
        ("EVCAT-002", "Vulnerability Scan Results", "Automated scanner outputs", "Confidential"),
        ("EVCAT-003", "Audit Workpapers", "Supporting audit documentation", "Confidential"),
        ("EVCAT-004", "System Configurations", "Captured configs during audit", "Confidential"),
        ("EVCAT-005", "Access Logs", "Records of auditor activity", "Internal"),
        ("EVCAT-006", "Interview Notes", "Documented conversations", "Confidential"),
        ("EVCAT-007", "Screenshots/Recordings", "Visual audit evidence", "Confidential"),
        ("EVCAT-008", "Final Audit Reports", "Published audit reports", "Confidential"),
        ("EVCAT-009", "Remediation Evidence", "Proof of fix implementation", "Internal"),
        ("EVCAT-010", "Management Response", "Management's reply to findings", "Internal"),
    ]

    for row_idx, (cat_id, cat_name, desc, classification) in enumerate(categories, 5):
        ws.cell(row=row_idx, column=1, value=cat_id)
        ws.cell(row=row_idx, column=2, value=cat_name)
        ws.cell(row=row_idx, column=3, value=desc)
        ws.cell(row=row_idx, column=4, value=classification)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Additional empty rows
    for row in range(15, 35):
        ws.cell(row=row, column=1, value=f"EVCAT-{str(row-4).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics
    ws.cell(row=37, column=1, value="PROTECTION STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Evidence Categories", "=COUNTA(B5:B34)"),
        ("Restricted Categories", '=COUNTIF(D5:D34,"Restricted")'),
        ("Confidential Categories", '=COUNTIF(D5:D34,"Confidential")'),
        ("Encryption at Rest (AES-256)", '=COUNTIF(H5:H34,"AES-256")'),
        ("Encryption in Transit (TLS 1.3)", '=COUNTIF(I5:I34,"TLS 1.3")'),
        ("Chain of Custody Required", '=COUNTIF(P5:P34,"Yes")'),
        ("Legal Hold Applicable", '=COUNTIF(R5:R34,"Yes")'),
    ]

    row = 39
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [12, 30, 50, 18, 40, 35, 35, 18, 18, 25, 40, 18, 25, 25, 25, 12, 40, 12, 25, 25, 12, 25, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_summary_dashboard_sheet(ws):
    """Create the Summary Dashboard sheet."""
    ws.title = "Summary_Dashboard"

    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="AUDIT ACTIVITY MANAGEMENT - SUMMARY DASHBOARD")
    title_cell.font = Font(bold=True, size=16, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Executive overview of audit governance compliance")
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.cell(row=3, column=1, value=f"Report Date: {GENERATED_DATE}")

    # Overall Compliance Summary
    ws.cell(row=5, column=1, value="OVERALL COMPLIANCE SUMMARY").font = Font(bold=True, size=12)

    headers = ["Metric", "Value", "Target", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    metrics = [
        ("Audit Authorization Rate", "[%]", "100%"),
        ("Tool Authorization Rate", "[%]", "100%"),
        ("Access Approval Rate", "[%]", "100%"),
        ("Access Revocation Timeliness", "[%]", "100%"),
        ("NDA Coverage", "[%]", "100%"),
        ("Access Logging Coverage", "[%]", "100%"),
        ("Critical System Mitigation Coverage", "[%]", "100%"),
        ("Evidence Encryption Rate", "[%]", "100%"),
    ]

    row = 7
    for metric, value, target in metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        cell_val = ws.cell(row=row, column=2, value=value)
        cell_val.fill = INPUT_FILL
        cell_val.border = THIN_BORDER
        ws.cell(row=row, column=3, value=target).border = THIN_BORDER
        cell_status = ws.cell(row=row, column=4)
        cell_status.fill = INPUT_FILL
        cell_status.border = THIN_BORDER
        row += 1

    # Status validation
    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant"')
    ws.add_data_validation(status_dv)
    status_dv.add('D7:D15')

    # Audit Activity Summary
    ws.cell(row=17, column=1, value="AUDIT ACTIVITY SUMMARY").font = Font(bold=True, size=12)

    activity_headers = ["Metric", "Count", "Status"]
    for col, header in enumerate(activity_headers, 1):
        cell = ws.cell(row=18, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    activity_items = [
        ("Total Audits This Period", "[#]"),
        ("Completed Audits", "[#]"),
        ("In Progress Audits", "[#]"),
        ("Planned Audits", "[#]"),
        ("Total Findings", "[#]"),
        ("Critical/High Findings", "[#]"),
        ("Open Follow-ups", "[#]"),
    ]

    row = 19
    for metric, count in activity_items:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        cell_count = ws.cell(row=row, column=2, value=count)
        cell_count.fill = INPUT_FILL
        cell_count.border = THIN_BORDER
        cell_status = ws.cell(row=row, column=3)
        cell_status.fill = INPUT_FILL
        cell_status.border = THIN_BORDER
        row += 1

    # Critical Findings
    ws.cell(row=28, column=1, value="CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION").font = Font(bold=True, size=12, color="C00000")

    findings_headers = ["Category", "Finding", "Count", "Severity", "Action Required"]
    for col, header in enumerate(findings_headers, 1):
        cell = ws.cell(row=29, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    findings = [
        ("Unauthorized Audits", "Audits without approval", "[#]", "Critical", "Immediate"),
        ("Unauthorized Tools", "Tools without authorization", "[#]", "Critical", "Immediate"),
        ("Unapproved Access", "Auditor access without approval", "[#]", "High", "Urgent"),
        ("Active Access Past End Date", "Access not revoked", "[#]", "High", "Urgent"),
        ("Missing Mitigation Plans", "Critical systems without plans", "[#]", "High", "Priority"),
        ("Unencrypted Evidence", "Sensitive evidence unprotected", "[#]", "Medium", "Plan"),
        ("Overdue Reviews", "Reviews past due", "[#]", "Medium", "Plan"),
    ]

    row = 30
    for cat, finding, count, severity, action in findings:
        ws.cell(row=row, column=1, value=cat).border = THIN_BORDER
        ws.cell(row=row, column=2, value=finding).border = THIN_BORDER
        cell_count = ws.cell(row=row, column=3, value=count)
        cell_count.fill = INPUT_FILL
        cell_count.border = THIN_BORDER
        ws.cell(row=row, column=4, value=severity).border = THIN_BORDER
        ws.cell(row=row, column=5, value=action).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [35, 40, 15, 15, 20, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_evidence_register_sheet(ws):
    """Create the Evidence Register sheet."""
    ws.title = "Evidence_Register"

    ws.merge_cells('A1:P1')
    title_cell = ws.cell(row=1, column=1, value="EVIDENCE REGISTER")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Documentation supporting assessment findings (100 entries)")

    headers = [
        "Evidence_ID", "Evidence_Type", "Evidence_Title", "Description",
        "Related_Assessment_Area", "Related_Finding_Control", "Document_Location",
        "Date_Collected", "Collected_By", "Verification_Status", "Verified_By",
        "Verification_Date", "Retention_Period", "Expiration_Date",
        "Confidentiality", "Auditor_Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(type="list", formula1='"Policy Document,Authorization Record,Access Log,Mitigation Plan,Audit Report,Tool Documentation,Other"')
    ws.add_data_validation(type_dv)
    type_dv.add('B5:B104')

    area_dv = DataValidation(type="list", formula1='"Audit Activity,Tool Authorization,Access Tracking,Disruption Mitigation,Evidence Protection"')
    ws.add_data_validation(area_dv)
    area_dv.add('E5:E104')

    status_dv = DataValidation(type="list", formula1='"Verified,Pending,Not Verified"')
    ws.add_data_validation(status_dv)
    status_dv.add('J5:J104')

    retention_dv = DataValidation(type="list", formula1='"1 Year,2 Years,3 Years,Permanent"')
    ws.add_data_validation(retention_dv)
    retention_dv.add('M5:M104')

    conf_dv = DataValidation(type="list", formula1='"Public,Internal,Confidential,Restricted"')
    ws.add_data_validation(conf_dv)
    conf_dv.add('O5:O104')

    # Format input rows
    for row in range(5, 105):
        ws.cell(row=row, column=1, value=f"EV-{str(row-4).zfill(3)}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Summary Statistics
    ws.cell(row=107, column=1, value="EVIDENCE STATISTICS").font = Font(bold=True, size=12)

    summary_items = [
        ("Total Evidence Entries", "=COUNTA(B5:B104)"),
        ("Verified Evidence", '=COUNTIF(J5:J104,"Verified")'),
        ("Pending Verification", '=COUNTIF(J5:J104,"Pending")'),
        ("Evidence by Area - Audit Activity", '=COUNTIF(E5:E104,"Audit Activity")'),
        ("Evidence by Area - Tool Authorization", '=COUNTIF(E5:E104,"Tool Authorization")'),
        ("Evidence by Area - Access Tracking", '=COUNTIF(E5:E104,"Access Tracking")'),
        ("Evidence by Area - Disruption Mitigation", '=COUNTIF(E5:E104,"Disruption Mitigation")'),
        ("Evidence by Area - Evidence Protection", '=COUNTIF(E5:E104,"Evidence Protection")'),
    ]

    row = 109
    for label, formula in summary_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=formula).border = THIN_BORDER
        row += 1

    # Column widths
    widths = [15, 25, 40, 50, 25, 25, 40, 15, 25, 18, 25, 15, 18, 15, 18, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_approval_sign_off_sheet(ws):
    """Create the Approval Sign-Off sheet."""
    ws.title = "Approval_Sign_Off"

    ws.merge_cells('A1:D1')
    title_cell = ws.cell(row=1, column=1, value="ASSESSMENT APPROVAL & SIGN-OFF")
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Three-level approval workflow for audit activity management assessment")

    # Assessment Summary
    ws.cell(row=4, column=1, value="ASSESSMENT SUMMARY").font = Font(bold=True, size=12)

    summary_items = [
        ("Document ID:", DOCUMENT_ID),
        ("Assessment Type:", "Audit Activity Management Assessment"),
        ("Assessment Period:", "[USER INPUT]"),
        ("Overall Compliance:", "[Formula from Summary_Dashboard]"),
        ("Critical Findings:", "[Formula]"),
        ("High Findings:", "[Formula]"),
        ("Evidence Entries:", "[Formula]"),
        ("Assessment Status:", "[Dropdown]"),
    ]

    row = 6
    for label, value in summary_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value or "[Dropdown]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Status dropdown
    status_dv = DataValidation(type="list", formula1='"Final,Requires Action,Draft,Re-assessment Required"')
    ws.add_data_validation(status_dv)
    status_dv.add('B13')

    # Level 1: Technical Validation
    row += 2
    ws.cell(row=row, column=1, value="LEVEL 1: TECHNICAL VALIDATION").font = Font(bold=True, size=12)
    row += 2

    level1_items = [
        ("Validator Name:", "[USER INPUT]"),
        ("Role/Title:", "IT Security Manager"),
        ("Validation Date:", "[DD.MM.YYYY]"),
        ("Validation Notes:", "[Text area]"),
        ("Technical Accuracy:", "[Dropdown]"),
        ("Recommendation:", "[Dropdown]"),
        ("Signature:", "[USER INPUT]"),
    ]

    for label, value in level1_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value or "[DD.MM.YYYY]" in value or "[Text area]" in value or "[Dropdown]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Level 1 dropdowns
    tech_acc_dv = DataValidation(type="list", formula1='"Confirmed,Minor Issues,Major Issues"')
    ws.add_data_validation(tech_acc_dv)
    tech_acc_dv.add('B22')

    rec_dv = DataValidation(type="list", formula1='"Approve,Approve with Conditions,Reject"')
    ws.add_data_validation(rec_dv)
    rec_dv.add('B23')

    # Level 2: Audit Approval
    row += 2
    ws.cell(row=row, column=1, value="LEVEL 2: AUDIT APPROVAL").font = Font(bold=True, size=12)
    row += 2

    level2_items = [
        ("Approver Name:", "[USER INPUT]"),
        ("Role/Title:", "Chief Audit Executive / Internal Audit Manager"),
        ("Approval Date:", "[DD.MM.YYYY]"),
        ("Audit Assessment:", "[Dropdown]"),
        ("Approval Decision:", "[Dropdown]"),
        ("Conditions:", "[Text area]"),
        ("Signature:", "[USER INPUT]"),
    ]

    for label, value in level2_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value or "[DD.MM.YYYY]" in value or "[Text area]" in value or "[Dropdown]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Level 2 dropdowns
    audit_assess_dv = DataValidation(type="list", formula1='"Acceptable,Acceptable with Conditions,Unacceptable"')
    ws.add_data_validation(audit_assess_dv)
    audit_assess_dv.add('B31')

    app_dec_dv = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected"')
    ws.add_data_validation(app_dec_dv)
    app_dec_dv.add('B32')

    # Level 3: Executive Confirmation
    row += 2
    ws.cell(row=row, column=1, value="LEVEL 3: EXECUTIVE CONFIRMATION").font = Font(bold=True, size=12)
    row += 2

    level3_items = [
        ("Approver Name:", "[USER INPUT]"),
        ("Role/Title:", "CISO / CRO"),
        ("Approval Date:", "[DD.MM.YYYY]"),
        ("Risk Acceptance:", "[Dropdown]"),
        ("Approval Decision:", "[Dropdown]"),
        ("Conditions:", "[Text area]"),
        ("Signature:", "[USER INPUT]"),
    ]

    for label, value in level3_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value or "[DD.MM.YYYY]" in value or "[Text area]" in value or "[Dropdown]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Level 3 dropdowns
    risk_dv = DataValidation(type="list", formula1='"Risks Acceptable,Conditional,Unacceptable"')
    ws.add_data_validation(risk_dv)
    risk_dv.add('B42')
    app_dec_dv.add('B43')

    # Next Review
    row += 2
    ws.cell(row=row, column=1, value="NEXT REVIEW").font = Font(bold=True, size=12)
    row += 2

    next_items = [
        ("Next Review Date:", "[Formula: Approval Date + 365 days]"),
        ("Review Responsibility:", "[USER INPUT]"),
        ("Focus Areas:", "[USER INPUT]"),
    ]

    for label, value in next_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if "[USER INPUT]" in value:
            cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30


def generate_workbook():
    """Generate the complete assessment workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Create all sheets
    create_instructions_sheet(wb.create_sheet())
    create_audit_activity_register_sheet(wb.create_sheet())
    create_audit_tool_authorization_sheet(wb.create_sheet())
    create_audit_access_tracking_sheet(wb.create_sheet())
    create_disruption_mitigation_sheet(wb.create_sheet())
    create_audit_evidence_protection_sheet(wb.create_sheet())
    create_summary_dashboard_sheet(wb.create_sheet())
    create_evidence_register_sheet(wb.create_sheet())
    create_approval_sign_off_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - INITIAL CREATION
# QA_TOOL: Claude Code
# CHANGES: Initial generator creation for A.8.33-34 Testing and Audit Protection
# =============================================================================
