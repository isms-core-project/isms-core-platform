#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.27.3 - Secure Architecture Pattern Catalogue Assessment Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.27: Secure System Architecture and Engineering
Assessment Domain 3 of 4: Secure Architecture Pattern Catalogue

--------------------------------------------------------------------------------
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.27.3"
WORKBOOK_NAME = "Secure Architecture Pattern Catalogue"
CONTROL_ID = "A.8.27"
CONTROL_NAME = "Secure System Architecture and Engineering Principles"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

CONFIG = {
    'organization': '[Organization]',
    'colors': {
        'header_bg': '1F4E79', 'header_text': 'FFFFFF', 'subheader_bg': '2E75B6',
        'input_cell': 'E2EFDA', 'formula_cell': 'FFF2CC',
    },
}

def get_styles():
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    return {
        'header': {'font': Font(bold=True, color=CONFIG['colors']['header_text'], size=12),
                   'fill': PatternFill(start_color=CONFIG['colors']['header_bg'], end_color=CONFIG['colors']['header_bg'], fill_type='solid'),
                   'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True), 'border': thin_border},
        'subheader': {'font': Font(bold=True, color=CONFIG['colors']['header_text'], size=11),
                      'fill': PatternFill(start_color=CONFIG['colors']['subheader_bg'], end_color=CONFIG['colors']['subheader_bg'], fill_type='solid'),
                      'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True), 'border': thin_border},
        'input': {'font': Font(size=11), 'fill': PatternFill(start_color=CONFIG['colors']['input_cell'], end_color=CONFIG['colors']['input_cell'], fill_type='solid'),
                  'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True), 'border': thin_border},
        'formula': {'font': Font(size=11), 'fill': PatternFill(start_color=CONFIG['colors']['formula_cell'], end_color=CONFIG['colors']['formula_cell'], fill_type='solid'),
                    'alignment': Alignment(horizontal='center', vertical='center'), 'border': thin_border},
        'normal': {'font': Font(size=11), 'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True), 'border': thin_border},
    }

def apply_style(cell, style_dict):
    for attr in ['font', 'fill', 'alignment', 'border']:
        if attr in style_dict:
            setattr(cell, attr, style_dict[attr])

def create_instructions_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:H1')
    ws['A1'] = f"ISMS-IMP-A.8.27.3 - Secure Architecture Pattern Catalogue Assessment"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    ws['A3'] = "Document Information"
    apply_style(ws['A3'], styles['subheader'])
    ws.merge_cells('A3:H3')

    info = [('Document ID', DOCUMENT_ID), ('Control Reference', CONTROL_REF),
            ('Purpose', 'Evaluate secure architecture pattern catalogue maturity'), ('Generated Date', GENERATED_DATE)]
    row = 4
    for label, value in info:
        ws[f'A{row}'], ws[f'B{row}'] = label, value
        apply_style(ws[f'A{row}'], styles['normal'])
        apply_style(ws[f'B{row}'], styles['input'])
        ws.merge_cells(f'B{row}:H{row}')
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80


def create_pattern_inventory_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:J1')
    ws['A1'] = "Secure Architecture Patterns - Pattern Inventory"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Pat-ID', 'Category', 'Name', 'Version', 'Status', 'Owner', 'LastReview', 'NextReview', 'DocumentRef', 'Notes']
    widths = [15, 20, 35, 10, 15, 20, 12, 12, 25, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    category_dv = DataValidation(type='list', formula1='"Authentication,Authorisation,DataProtection,NetworkSecurity,Integration,Cloud,LoggingMonitoring,Identity,Container,Serverless"', allow_blank=True)
    ws.add_data_validation(category_dv)

    status_dv = DataValidation(type='list', formula1='"Approved,Draft,Deprecated,Under Review"', allow_blank=True)
    ws.add_data_validation(status_dv)

    patterns = [
        ('PAT-AUTH-001', 'Authentication', 'SSO Integration Pattern'),
        ('PAT-AUTH-002', 'Authentication', 'MFA Implementation Pattern'),
        ('PAT-AUTH-003', 'Authentication', 'Service Authentication Pattern'),
        ('PAT-AUTH-004', 'Authentication', 'Token Management Pattern'),
        ('PAT-AUTHZ-001', 'Authorisation', 'RBAC Implementation Pattern'),
        ('PAT-AUTHZ-002', 'Authorisation', 'ABAC Implementation Pattern'),
        ('PAT-AUTHZ-003', 'Authorisation', 'API Authorisation Pattern'),
        ('PAT-AUTHZ-004', 'Authorisation', 'Delegated Administration Pattern'),
        ('PAT-DATA-001', 'DataProtection', 'Encryption at Rest Pattern'),
        ('PAT-DATA-002', 'DataProtection', 'Encryption in Transit Pattern'),
        ('PAT-DATA-003', 'DataProtection', 'Key Management Pattern'),
        ('PAT-DATA-004', 'DataProtection', 'Tokenisation Pattern'),
        ('PAT-NET-001', 'NetworkSecurity', 'DMZ Architecture Pattern'),
        ('PAT-NET-002', 'NetworkSecurity', 'Micro-segmentation Pattern'),
        ('PAT-NET-003', 'NetworkSecurity', 'API Gateway Pattern'),
        ('PAT-NET-004', 'NetworkSecurity', 'Zero Trust Network Pattern'),
        ('PAT-INT-001', 'Integration', 'Secure REST API Pattern'),
        ('PAT-INT-002', 'Integration', 'Message Queue Security Pattern'),
        ('PAT-INT-003', 'Integration', 'Event-Driven Security Pattern'),
        ('PAT-INT-004', 'Integration', 'Service Mesh Pattern'),
        ('PAT-CLD-001', 'Cloud', 'Landing Zone Architecture'),
        ('PAT-CLD-002', 'Cloud', 'Workload Isolation Pattern'),
        ('PAT-CLD-003', 'Cloud', 'Serverless Security Pattern'),
        ('PAT-CLD-004', 'Cloud', 'Container Security Pattern'),
        ('PAT-LOG-001', 'LoggingMonitoring', 'Centralised Logging Pattern'),
        ('PAT-LOG-002', 'LoggingMonitoring', 'SIEM Integration Pattern'),
        ('PAT-LOG-003', 'LoggingMonitoring', 'Audit Trail Pattern'),
        ('PAT-LOG-004', 'LoggingMonitoring', 'Log Protection Pattern'),
        ('PAT-IDM-001', 'Identity', 'Identity Lifecycle Pattern'),
        ('PAT-IDM-002', 'Identity', 'Privileged Access Management Pattern'),
    ]

    for row, (pat_id, category, name) in enumerate(patterns, 4):
        ws.cell(row=row, column=1, value=pat_id)
        ws.cell(row=row, column=2, value=category)
        ws.cell(row=row, column=3, value=name)
        for col in range(1, 11):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 3 else styles['normal'])

    category_dv.add(f'B4:B{3 + len(patterns)}')
    status_dv.add(f'E4:E{3 + len(patterns)}')
    ws.freeze_panes = 'A4'


def create_pattern_quality_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:E1')
    ws['A1'] = "Secure Architecture Patterns - Documentation Quality"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Pat-ID', 'Element', 'Present', 'Quality', 'Notes']
    widths = [15, 25, 10, 10, 40]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    present_dv = DataValidation(type='list', formula1='"Yes,Partial,No"', allow_blank=True)
    ws.add_data_validation(present_dv)

    rating_dv = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(rating_dv)

    elements = ['ProblemStatement', 'Context', 'Solution', 'SecurityRationale', 'Implementation',
                'Example', 'AntiPatterns', 'RelatedPatterns', 'ComplianceMapping', 'ThreatModel']

    row = 4
    for element in elements:
        ws.cell(row=row, column=2, value=element)
        for col in range(1, 6):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col != 2 else styles['normal'])
        row += 1

    present_dv.add(f'C4:C{row-1}')
    rating_dv.add(f'D4:D{row-1}')
    ws.freeze_panes = 'A4'


def create_adoption_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:G1')
    ws['A1'] = "Secure Architecture Patterns - Adoption Tracking"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Pat-ID', 'ProjectCount', 'TotalApplicable', 'AdoptionRate', 'Trend', 'Barriers', 'Action']
    widths = [15, 15, 15, 15, 12, 35, 30]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    trend_dv = DataValidation(type='list', formula1='"Increasing,Stable,Decreasing"', allow_blank=True)
    ws.add_data_validation(trend_dv)

    for row in range(4, 34):
        ws.cell(row=row, column=4, value=f'=IF(C{row}>0,B{row}/C{row},0)')
        for col in range(1, 8):
            apply_style(ws.cell(row=row, column=col), styles['formula'] if col == 4 else styles['input'])

    trend_dv.add('E4:E33')
    ws.freeze_panes = 'A4'


def create_governance_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:G1')
    ws['A1'] = "Secure Architecture Patterns - Governance Assessment"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Gov-ID', 'Requirement', 'Status', 'Evidence', 'Gap', 'Owner', 'Notes']
    widths = [10, 40, 15, 30, 30, 20, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    status_dv = DataValidation(type='list', formula1='"Implemented,Partial,Not Implemented"', allow_blank=True)
    ws.add_data_validation(status_dv)

    governance_reqs = [
        ('GOV-001', 'Pattern ownership assigned for each pattern'),
        ('GOV-002', 'Annual pattern review cycle established'),
        ('GOV-003', 'Patterns linked to threat models'),
        ('GOV-004', 'Pattern versioning and change control'),
        ('GOV-005', 'Pattern update process documented'),
        ('GOV-006', 'Pattern deprecation process defined'),
        ('GOV-007', 'Pattern changes communicated to stakeholders'),
        ('GOV-008', 'Pattern training available'),
        ('GOV-009', 'Exception process for pattern deviations'),
        ('GOV-010', 'Pattern effectiveness measured'),
        ('GOV-011', 'Pattern catalogue accessible to all architects'),
        ('GOV-012', 'New pattern proposal process defined'),
        ('GOV-013', 'Pattern review board established'),
        ('GOV-014', 'Integration with architecture review process'),
        ('GOV-015', 'Pattern compliance tracked in projects'),
    ]

    for row, (gov_id, requirement) in enumerate(governance_reqs, 4):
        ws.cell(row=row, column=1, value=gov_id)
        ws.cell(row=row, column=2, value=requirement)
        for col in range(1, 8):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 2 else styles['normal'])

    status_dv.add(f'C4:C{3 + len(governance_reqs)}')
    ws.freeze_panes = 'A4'


def create_deviations_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:J1')
    ws['A1'] = "Secure Architecture Patterns - Deviation Tracking"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Dev-ID', 'Pattern', 'Project', 'Category', 'Justification', 'Approved', 'ApprovedBy', 'Compensating', 'Expiry', 'Status']
    widths = [10, 15, 25, 20, 40, 10, 20, 35, 12, 12]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    category_dv = DataValidation(type='list', formula1='"Technical Constraint,Legacy System,Vendor Limitation,Cost/Time,Performance"', allow_blank=True)
    ws.add_data_validation(category_dv)

    approved_dv = DataValidation(type='list', formula1='"Yes,No,Pending"', allow_blank=True)
    ws.add_data_validation(approved_dv)

    status_dv = DataValidation(type='list', formula1='"Active,Closed,Expired"', allow_blank=True)
    ws.add_data_validation(status_dv)

    for row in range(4, 24):
        ws.cell(row=row, column=1, value=f'DEV-{row-3:03d}')
        for col in range(1, 11):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 1 else styles['normal'])

    category_dv.add('D4:D23')
    approved_dv.add('F4:F23')
    status_dv.add('J4:J23')
    ws.freeze_panes = 'A4'


def create_effectiveness_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:G1')
    ws['A1'] = "Secure Architecture Patterns - Effectiveness Metrics"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Pat-ID', 'SecurityIncidents', 'VulnFindings', 'AuditFindings', 'UserFeedback', 'Effectiveness', 'Notes']
    widths = [15, 15, 20, 15, 12, 12, 30]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    effectiveness_dv = DataValidation(type='list', formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(effectiveness_dv)

    for row in range(4, 34):
        for col in range(1, 8):
            apply_style(ws.cell(row=row, column=col), styles['input'])

    effectiveness_dv.add('F4:F33')
    ws.freeze_panes = 'A4'


def create_compliance_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:G1')
    ws['A1'] = "Secure Architecture Patterns - Policy Compliance"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Comp-ID', 'Requirement', 'Source', 'Compliant', 'Evidence', 'Score', 'Notes']
    widths = [10, 40, 20, 12, 35, 10, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    compliant_dv = DataValidation(type='list', formula1='"Yes,Partial,No"', allow_blank=True)
    ws.add_data_validation(compliant_dv)

    compliance_reqs = [
        ('COMP-001', 'Secure pattern catalogue maintained', 'POL-A.8.27 §2.2.2'),
        ('COMP-002', 'Pattern documentation with security rationale', 'POL-A.8.27 §2.2.2'),
        ('COMP-003', 'Annual pattern review conducted', 'POL-A.8.27 §2.2.2'),
        ('COMP-004', 'Deviations require Security Architect approval', 'POL-A.8.27 §2.2.2'),
        ('COMP-005', 'New patterns validated through threat modelling', 'POL-A.8.27 §2.2.2'),
        ('COMP-006', 'Authentication patterns defined', 'POL-A.8.27 §2.2.2'),
        ('COMP-007', 'Authorisation patterns defined', 'POL-A.8.27 §2.2.2'),
        ('COMP-008', 'Data protection patterns defined', 'POL-A.8.27 §2.2.2'),
        ('COMP-009', 'Network security patterns defined', 'POL-A.8.27 §2.2.2'),
        ('COMP-010', 'Integration patterns defined', 'POL-A.8.27 §2.2.2'),
        ('COMP-011', 'Cloud patterns defined', 'POL-A.8.27 §2.2.2'),
        ('COMP-012', 'Pattern adoption tracked', 'POL-A.8.27 §4'),
        ('COMP-013', 'Pattern effectiveness measured', 'POL-A.8.27 §4'),
        ('COMP-014', 'Deviation tracking maintained', 'POL-A.8.27 §4'),
        ('COMP-015', 'Pattern training provided', 'POL-A.8.27 §4'),
    ]

    for row, (comp_id, requirement, source) in enumerate(compliance_reqs, 4):
        ws.cell(row=row, column=1, value=comp_id)
        ws.cell(row=row, column=2, value=requirement)
        ws.cell(row=row, column=3, value=source)
        ws.cell(row=row, column=6, value=f'=IF(D{row}="Yes",100,IF(D{row}="Partial",50,0))')
        for col in range(1, 8):
            if col == 6:
                apply_style(ws.cell(row=row, column=col), styles['formula'])
            elif col > 3:
                apply_style(ws.cell(row=row, column=col), styles['input'])
            else:
                apply_style(ws.cell(row=row, column=col), styles['normal'])

    compliant_dv.add(f'D4:D{3 + len(compliance_reqs)}')

    last_row = 3 + len(compliance_reqs) + 2
    ws.cell(row=last_row, column=5, value="Overall Compliance:")
    ws.cell(row=last_row, column=6, value=f'=AVERAGE(F4:F{3 + len(compliance_reqs)})')
    apply_style(ws.cell(row=last_row, column=5), styles['subheader'])
    apply_style(ws.cell(row=last_row, column=6), styles['formula'])

    ws.freeze_panes = 'A4'


def create_gap_register_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:J1')
    ws['A1'] = "Secure Architecture Patterns - Gap Register"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Gap-ID', 'Source', 'Description', 'Risk', 'Remediation', 'Owner', 'Due Date', 'Status', 'Closure Date', 'Notes']
    widths = [10, 15, 40, 10, 40, 20, 12, 12, 12, 25]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    risk_dv = DataValidation(type='list', formula1='"High,Medium,Low"', allow_blank=True)
    status_dv = DataValidation(type='list', formula1='"Open,In Progress,Closed"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    ws.add_data_validation(status_dv)

    for row in range(4, 24):
        ws.cell(row=row, column=1, value=f'GAP-{row-3:03d}')
        for col in range(1, 11):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 1 else styles['normal'])

    risk_dv.add('D4:D23')
    status_dv.add('H4:H23')
    ws.freeze_panes = 'A4'


def create_dashboard_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:F1')
    ws['A1'] = "Secure Architecture Patterns - Assessment Dashboard"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    ws['A3'] = "Assessment Summary"
    apply_style(ws['A3'], styles['subheader'])
    ws.merge_cells('A3:F3')

    summary = [
        ('Assessment Date', GENERATED_DATE),
        ('Organization', CONFIG['organization']),
        ('Control Reference', CONTROL_REF),
        ('Total Patterns', '=COUNTA(PatternInventory!A4:A100)'),
        ('Overall Compliance Score', '=Compliance!F20'),
    ]

    row = 4
    for label, value in summary:
        ws[f'A{row}'], ws[f'B{row}'] = label, value
        apply_style(ws[f'A{row}'], styles['normal'])
        apply_style(ws[f'B{row}'], styles['formula'] if '=' in str(value) else styles['input'])
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def main():
    logger.info("=" * 80)
    logger.info(f"ISMS Control {CONTROL_ID} - {WORKBOOK_NAME} Generator")
    logger.info("=" * 80)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        ("Instructions", create_instructions_sheet),
        ("PatternInventory", create_pattern_inventory_sheet),
        ("PatternQuality", create_pattern_quality_sheet),
        ("Adoption", create_adoption_sheet),
        ("Governance", create_governance_sheet),
        ("Deviations", create_deviations_sheet),
        ("Effectiveness", create_effectiveness_sheet),
        ("Compliance", create_compliance_sheet),
        ("GapRegister", create_gap_register_sheet),
        ("Dashboard", create_dashboard_sheet),
    ]

    for sheet_name, create_func in sheets:
        ws = wb.create_sheet(title=sheet_name)
        create_func(ws)
        logger.info(f"  ✓ Created sheet: {sheet_name}")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, '..', '90_workbooks')
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, OUTPUT_FILENAME)

    wb.save(output_path)
    logger.info(f"Workbook saved: {output_path}")
    logger.info("Generation complete!")


if __name__ == "__main__":
    main()


# =============================================================================
# QA_VERIFIED: 2026-02-02
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation with standard structure
# =============================================================================
