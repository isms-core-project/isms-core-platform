#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.27.5 - SSE Compliance Dashboard Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.27: Secure System Architecture and Engineering
Assessment Domain 5 of 5: Consolidated SSE Compliance Dashboard

Consolidates data from all 4 SSE assessment domains for executive reporting.

--------------------------------------------------------------------------------
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.27.5"
WORKBOOK_NAME = "SSE Compliance Dashboard"
CONTROL_ID = "A.8.27"
CONTROL_NAME = "Secure System Architecture and Engineering Principles"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

CONFIG = {
    'organization': '[Organization]',
    'colors': {
        'header_bg': '1F4E79', 'header_text': 'FFFFFF', 'subheader_bg': '2E75B6',
        'input_cell': 'E2EFDA', 'formula_cell': 'FFF2CC',
        'green': '2ECC71', 'amber': 'F39C12', 'red': 'E74C3C',
    },
}

def get_styles():
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    return {
        'header': {'font': Font(bold=True, color=CONFIG['colors']['header_text'], size=12),
                   'fill': PatternFill(start_color=CONFIG['colors']['header_bg'], end_color=CONFIG['colors']['header_bg'], fill_type='solid'),
                   'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True), 'border': thin_border},
        'subheader': {'font': Font(bold=True, color=CONFIG['colors']['header_text'], size=11),
                      'fill': PatternFill(start_color=CONFIG['colors']['subheader_bg'], end_color=CONFIG['colors']['subheader_bg'], fill_type='solid'),
                      'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True), 'border': thin_border},
        'input': {'font': Font(size=11), 'fill': PatternFill(start_color=CONFIG['colors']['input_cell'], end_color=CONFIG['colors']['input_cell'], fill_type='solid'),
                  'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True), 'border': thin_border},
        'formula': {'font': Font(size=11), 'fill': PatternFill(start_color=CONFIG['colors']['formula_cell'], end_color=CONFIG['colors']['formula_cell'], fill_type='solid'),
                    'alignment': Alignment(horizontal='center', vertical='center'), 'border': thin_border},
        'normal': {'font': Font(size=11), 'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True), 'border': thin_border},
        'big_number': {'font': Font(bold=True, size=36, color=CONFIG['colors']['header_bg']),
                       'alignment': Alignment(horizontal='center', vertical='center')},
        'title': {'font': Font(bold=True, size=18, color=CONFIG['colors']['header_bg']),
                  'alignment': Alignment(horizontal='center', vertical='center')},
    }

def apply_style(cell, style_dict):
    for attr in ['font', 'fill', 'alignment', 'border']:
        if attr in style_dict:
            setattr(cell, attr, style_dict[attr])


def create_instructions_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:H1')
    ws['A1'] = f"ISMS-IMP-A.8.27.5 - SSE Compliance Dashboard"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    ws['A3'] = "Dashboard Overview"
    apply_style(ws['A3'], styles['subheader'])
    ws.merge_cells('A3:H3')

    info = [
        ('Document ID', DOCUMENT_ID),
        ('Control Reference', CONTROL_REF),
        ('Purpose', 'Consolidated SSE compliance dashboard for executive reporting'),
        ('Data Sources', 'A.8.27.1-4 Assessment Workbooks'),
        ('Update Frequency', 'Quarterly (after domain assessments)'),
        ('Generated Date', GENERATED_DATE),
    ]
    row = 4
    for label, value in info:
        ws[f'A{row}'], ws[f'B{row}'] = label, value
        apply_style(ws[f'A{row}'], styles['normal'])
        apply_style(ws[f'B{row}'], styles['input'])
        ws.merge_cells(f'B{row}:H{row}')
        row += 1

    row += 1
    ws[f'A{row}'] = "Dashboard Sheets"
    apply_style(ws[f'A{row}'], styles['subheader'])
    ws.merge_cells(f'A{row}:H{row}')
    row += 1

    sheets_info = [
        ('ExecutiveSummary', 'High-level SSE status for leadership'),
        ('DomainScores', 'Compliance scores by assessment domain'),
        ('ZeroTrustRadar', 'Zero Trust pillar maturity data'),
        ('GapConsolidation', 'All SSE gaps from all domains'),
        ('TrendAnalysis', 'Historical trend data'),
        ('AuditEvidence', 'Evidence inventory for audits'),
        ('ActionTracker', 'Remediation action tracking'),
    ]
    for sheet, desc in sheets_info:
        ws[f'A{row}'], ws[f'B{row}'] = sheet, desc
        apply_style(ws[f'A{row}'], styles['normal'])
        apply_style(ws[f'B{row}'], styles['normal'])
        ws.merge_cells(f'B{row}:H{row}')
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60


def create_executive_summary_sheet(ws):
    styles = get_styles()

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Secure Systems Engineering - Executive Dashboard"
    apply_style(ws['A1'], styles['title'])
    ws.row_dimensions[1].height = 40

    ws['A2'] = f"Report Date: {GENERATED_DATE}"
    ws['A2'].font = Font(size=11, italic=True)

    # Overall Score Section
    ws.merge_cells('A4:C4')
    ws['A4'] = "Overall SSE Compliance Score"
    apply_style(ws['A4'], styles['subheader'])

    ws.merge_cells('A5:C7')
    ws['A5'] = '=AVERAGE(DomainScores!D4:D7)'
    apply_style(ws['A5'], styles['big_number'])
    ws.row_dimensions[5].height = 50
    ws.row_dimensions[6].height = 30
    ws.row_dimensions[7].height = 30

    # Domain Status Table
    ws['A9'] = "Domain Status Summary"
    apply_style(ws['A9'], styles['subheader'])
    ws.merge_cells('A9:H9')

    headers = ['Domain', 'Name', 'Score', 'Status', 'Gaps', 'Trend']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        apply_style(cell, styles['header'])

    domains = [
        ('A.8.27.1', 'Architecture Review Process', '=DomainScores!D4'),
        ('A.8.27.2', 'Threat Modelling Methodology', '=DomainScores!D5'),
        ('A.8.27.3', 'Secure Architecture Patterns', '=DomainScores!D6'),
        ('A.8.27.4', 'Zero Trust Implementation', '=DomainScores!D7'),
    ]

    for row, (domain, name, score_ref) in enumerate(domains, 11):
        ws.cell(row=row, column=1, value=domain)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=score_ref)
        for col in range(1, 7):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 2 else styles['normal'])

    # Key Metrics
    ws['A16'] = "Key Metrics"
    apply_style(ws['A16'], styles['subheader'])
    ws.merge_cells('A16:H16')

    metrics = [
        ('Total Open Gaps', '=COUNTIF(GapConsolidation!H:H,"Open")'),
        ('High Risk Gaps', '=COUNTIF(GapConsolidation!D:D,"High")'),
        ('Overdue Actions', '=COUNTIF(ActionTracker!F:F,"Overdue")'),
        ('ZT Average Maturity', '=AVERAGE(ZeroTrustRadar!C4:C10)'),
    ]

    row = 17
    for metric, formula in metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=formula)
        apply_style(ws.cell(row=row, column=1), styles['normal'])
        apply_style(ws.cell(row=row, column=2), styles['formula'])
        row += 1

    # Next Steps
    ws['A22'] = "Priority Actions for Next Quarter"
    apply_style(ws['A22'], styles['subheader'])
    ws.merge_cells('A22:H22')

    for row in range(23, 28):
        ws.cell(row=row, column=1, value=f'{row-22}.')
        apply_style(ws.cell(row=row, column=1), styles['normal'])
        ws.merge_cells(f'B{row}:H{row}')
        apply_style(ws.cell(row=row, column=2), styles['input'])

    # Column widths
    widths = [15, 35, 12, 12, 10, 10, 15, 15]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_domain_scores_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:I1')
    ws['A1'] = "SSE Domain Compliance Scores"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Domain', 'Name', 'Assessment Date', 'Compliance Score', 'Gap Count', 'High Risk', 'Status', 'Trend', 'Next Assessment']
    widths = [12, 40, 15, 15, 12, 12, 10, 10, 15]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    status_dv = DataValidation(type='list', formula1='"Green,Amber,Red"', allow_blank=True)
    trend_dv = DataValidation(type='list', formula1='"↑,→,↓"', allow_blank=True)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(trend_dv)

    domains = [
        ('A.8.27.1', 'Security Architecture Review Process'),
        ('A.8.27.2', 'Threat Modelling Methodology'),
        ('A.8.27.3', 'Secure Architecture Pattern Catalogue'),
        ('A.8.27.4', 'Zero Trust Implementation'),
    ]

    for row, (domain, name) in enumerate(domains, 4):
        ws.cell(row=row, column=1, value=domain)
        ws.cell(row=row, column=2, value=name)
        for col in range(1, 10):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 2 else styles['normal'])

    status_dv.add('G4:G7')
    trend_dv.add('H4:H7')

    # Summary row
    ws.cell(row=9, column=1, value="OVERALL")
    ws.cell(row=9, column=4, value='=AVERAGE(D4:D7)')
    ws.cell(row=9, column=5, value='=SUM(E4:E7)')
    ws.cell(row=9, column=6, value='=SUM(F4:F7)')
    apply_style(ws.cell(row=9, column=1), styles['subheader'])
    for col in range(4, 7):
        apply_style(ws.cell(row=9, column=col), styles['formula'])

    ws.freeze_panes = 'A4'


def create_zero_trust_radar_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:G1')
    ws['A1'] = "Zero Trust Maturity - Radar Chart Data"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Pillar', 'Current Level', 'Current Score', 'Target Level', 'Target Score', 'Gap', 'Priority']
    widths = [15, 15, 12, 15, 12, 10, 10]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    level_dv = DataValidation(type='list', formula1='"Traditional,Initial,Advanced,Optimal"', allow_blank=True)
    score_dv = DataValidation(type='list', formula1='"1,2,3,4"', allow_blank=True)
    priority_dv = DataValidation(type='list', formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(level_dv)
    ws.add_data_validation(score_dv)
    ws.add_data_validation(priority_dv)

    pillars = ['Identity', 'Device', 'Network', 'Workload', 'Data', 'Visibility', 'Automation']

    for row, pillar in enumerate(pillars, 4):
        ws.cell(row=row, column=1, value=pillar)
        ws.cell(row=row, column=6, value=f'=E{row}-C{row}')
        for col in range(1, 8):
            if col == 6:
                apply_style(ws.cell(row=row, column=col), styles['formula'])
            elif col > 1:
                apply_style(ws.cell(row=row, column=col), styles['input'])
            else:
                apply_style(ws.cell(row=row, column=col), styles['normal'])

    level_dv.add('B4:B10')
    level_dv.add('D4:D10')
    score_dv.add('C4:C10')
    score_dv.add('E4:E10')
    priority_dv.add('G4:G10')

    # Average row
    ws.cell(row=12, column=1, value="AVERAGE")
    ws.cell(row=12, column=3, value='=AVERAGE(C4:C10)')
    ws.cell(row=12, column=5, value='=AVERAGE(E4:E10)')
    ws.cell(row=12, column=6, value='=E12-C12')
    apply_style(ws.cell(row=12, column=1), styles['subheader'])
    for col in [3, 5, 6]:
        apply_style(ws.cell(row=12, column=col), styles['formula'])

    ws.freeze_panes = 'A4'


def create_gap_consolidation_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:J1')
    ws['A1'] = "SSE Gap Consolidation"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Gap-ID', 'Source', 'Description', 'Risk', 'Remediation', 'Owner', 'Due Date', 'Status', 'Days Open', 'Overdue']
    widths = [10, 12, 45, 10, 40, 20, 12, 12, 10, 10]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    source_dv = DataValidation(type='list', formula1='"A.8.27.1,A.8.27.2,A.8.27.3,A.8.27.4"', allow_blank=True)
    risk_dv = DataValidation(type='list', formula1='"High,Medium,Low"', allow_blank=True)
    status_dv = DataValidation(type='list', formula1='"Open,In Progress,Closed"', allow_blank=True)
    ws.add_data_validation(source_dv)
    ws.add_data_validation(risk_dv)
    ws.add_data_validation(status_dv)

    for row in range(4, 54):
        ws.cell(row=row, column=1, value=f'SSE-GAP-{row-3:03d}')
        ws.cell(row=row, column=9, value=f'=IF(H{row}="Closed","",TODAY()-G{row})')
        ws.cell(row=row, column=10, value=f'=IF(AND(H{row}<>"Closed",G{row}<TODAY()),"OVERDUE","")')
        for col in range(1, 11):
            if col in [9, 10]:
                apply_style(ws.cell(row=row, column=col), styles['formula'])
            elif col > 1:
                apply_style(ws.cell(row=row, column=col), styles['input'])
            else:
                apply_style(ws.cell(row=row, column=col), styles['normal'])

    source_dv.add('B4:B53')
    risk_dv.add('D4:D53')
    status_dv.add('H4:H53')

    ws.freeze_panes = 'A4'


def create_trend_analysis_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:H1')
    ws['A1'] = "SSE Compliance Trend Analysis"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Period', 'Overall Score', 'Domain 1', 'Domain 2', 'Domain 3', 'Domain 4', 'Open Gaps', 'High Gaps Closed']
    widths = [12, 12, 12, 12, 12, 12, 12, 15]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    # Pre-populate 8 quarters (2 years)
    periods = ['Q1 2026', 'Q2 2026', 'Q3 2026', 'Q4 2026', 'Q1 2027', 'Q2 2027', 'Q3 2027', 'Q4 2027']

    for row, period in enumerate(periods, 4):
        ws.cell(row=row, column=1, value=period)
        for col in range(1, 9):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 1 else styles['normal'])

    ws.freeze_panes = 'A4'


def create_audit_evidence_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:G1')
    ws['A1'] = "SSE Audit Evidence Inventory"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Evidence-ID', 'Domain', 'Description', 'Location', 'Date', 'Status', 'Audit Use']
    widths = [12, 12, 45, 40, 12, 12, 15]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    domain_dv = DataValidation(type='list', formula1='"A.8.27.1,A.8.27.2,A.8.27.3,A.8.27.4,All"', allow_blank=True)
    status_dv = DataValidation(type='list', formula1='"Available,Pending,Missing"', allow_blank=True)
    audit_dv = DataValidation(type='list', formula1='"Stage 1,Stage 2,Both"', allow_blank=True)
    ws.add_data_validation(domain_dv)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(audit_dv)

    # Pre-populate evidence items
    evidence_items = [
        ('EV-001', 'All', 'ISMS-POL-A.8.27 Approved Policy'),
        ('EV-002', 'A.8.27.1', 'Architecture Review Procedures'),
        ('EV-003', 'A.8.27.1', 'Sample Architecture Reviews (3+)'),
        ('EV-004', 'A.8.27.2', 'Threat Modelling Methodology'),
        ('EV-005', 'A.8.27.2', 'Sample Threat Models (3+)'),
        ('EV-006', 'A.8.27.2', 'MITRE ATT&CK Coverage Matrix'),
        ('EV-007', 'A.8.27.3', 'Secure Pattern Catalogue'),
        ('EV-008', 'A.8.27.3', 'Pattern Adoption Metrics'),
        ('EV-009', 'A.8.27.4', 'Zero Trust Strategy Document'),
        ('EV-010', 'A.8.27.4', 'ZT Maturity Assessment Results'),
        ('EV-011', 'All', 'SSE Training Records'),
        ('EV-012', 'All', 'Quarterly Dashboard Reports'),
    ]

    for row, (ev_id, domain, desc) in enumerate(evidence_items, 4):
        ws.cell(row=row, column=1, value=ev_id)
        ws.cell(row=row, column=2, value=domain)
        ws.cell(row=row, column=3, value=desc)
        for col in range(1, 8):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 3 else styles['normal'])

    domain_dv.add(f'B4:B{3 + len(evidence_items) + 10}')
    status_dv.add(f'F4:F{3 + len(evidence_items) + 10}')
    audit_dv.add(f'G4:G{3 + len(evidence_items) + 10}')

    ws.freeze_panes = 'A4'


def create_action_tracker_sheet(ws):
    styles = get_styles()
    ws.merge_cells('A1:H1')
    ws['A1'] = "SSE Remediation Action Tracker"
    apply_style(ws['A1'], styles['header'])
    ws.row_dimensions[1].height = 30

    headers = ['Action-ID', 'Gap-ID', 'Action', 'Owner', 'Due Date', 'Status', 'Completion Date', 'Evidence']
    widths = [10, 15, 45, 20, 12, 15, 15, 35]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_style(cell, styles['header'])
        ws.column_dimensions[get_column_letter(col)].width = width

    status_dv = DataValidation(type='list', formula1='"Not Started,In Progress,Completed,Overdue"', allow_blank=True)
    ws.add_data_validation(status_dv)

    for row in range(4, 34):
        ws.cell(row=row, column=1, value=f'ACT-{row-3:03d}')
        for col in range(1, 9):
            apply_style(ws.cell(row=row, column=col), styles['input'] if col > 1 else styles['normal'])

    status_dv.add('F4:F33')

    ws.freeze_panes = 'A4'


def main():
    logger.info("=" * 80)
    logger.info(f"ISMS Control {CONTROL_ID} - {WORKBOOK_NAME} Generator")
    logger.info("=" * 80)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        ("Instructions", create_instructions_sheet),
        ("ExecutiveSummary", create_executive_summary_sheet),
        ("DomainScores", create_domain_scores_sheet),
        ("ZeroTrustRadar", create_zero_trust_radar_sheet),
        ("GapConsolidation", create_gap_consolidation_sheet),
        ("TrendAnalysis", create_trend_analysis_sheet),
        ("AuditEvidence", create_audit_evidence_sheet),
        ("ActionTracker", create_action_tracker_sheet),
    ]

    for sheet_name, create_func in sheets:
        ws = wb.create_sheet(title=sheet_name)
        create_func(ws)
        logger.info(f"  ✓ Created sheet: {sheet_name}")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, '..', '90_workbooks')
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, OUTPUT_FILENAME)

    wb.save(output_path)
    logger.info(f"Workbook saved: {output_path}")
    logger.info("Generation complete!")


if __name__ == "__main__":
    main()


# =============================================================================
# QA_VERIFIED: 2026-02-02
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation with standard structure
# =============================================================================
