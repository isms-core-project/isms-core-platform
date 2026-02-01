#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS A.8.13/14/5.30 - Assessment File Normalization Utility
================================================================================

ISO/IEC 27001:2022 Controls A.8.13 (Backup), A.8.14 (Redundancy), A.5.30 (ICT BC)
Quality Assurance Utility: File Preparation for Dashboard Linking

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific file naming conventions, assessment file locations,
and dashboard integration requirements.

Key customization areas:
1. Expected document IDs (match your assessment workbook identifiers)
2. File naming patterns (adapt to your organizational standards)
3. Source and output directory structures
4. Validation rules (customize based on your quality requirements)
5. Manifest format and content

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.13/14/5.30 BC/DR Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script normalizes BC/DR assessment file names to enable external workbook
formula linking in the ISMS-IMP-A.8.13.S5 BC/DR Compliance Dashboard. It scans
for completed assessments, validates document IDs, and creates normalized copies
with standardized names.

**Purpose:**
Prepares BC/DR assessment workbooks for dashboard consolidation by removing
date suffixes and version numbers from filenames, enabling stable external
workbook formula references that don't break when files are updated.

**Normalization Process:**
1. Scans source directory for Excel assessment files
2. Opens each file and reads Document ID from Instructions sheet
3. Validates Document ID matches expected BC/DR assessment identifiers
4. Copies file to output directory with normalized filename
5. Creates audit manifest documenting normalization

**File Name Transformation:**
Before normalization (date/version-specific names):
- ISMS_Assessment_Backup_Inventory_20250125.xlsx
- ISMS_Assessment_Redundancy_Analysis_v2.xlsx
- RPO_RTO_Compliance_Final.xlsx

After normalization (stable names for dashboard linking):
- ISMS-IMP-A.8.13.S1.xlsx
- ISMS-IMP-A.8.13.S2.xlsx
- ISMS-IMP-A.8.13.S3.xlsx

Dashboard external formulas reference normalized names, remaining stable across
file updates.

**Validation Performed:**
- Checks for presence of "Instructions" or "Instructions & Legend" sheet
- Locates "Document ID:" field in sheet (rows 3-25, column B)
- Validates Document ID matches one of:
  * ISMS-IMP-A.8.13.S1 (Backup Inventory)
  * ISMS-IMP-A.8.13.S2 (Redundancy Analysis)
  * ISMS-IMP-A.8.13.S3 (RPO/RTO Compliance)
  * ISMS-IMP-A.8.13.S4 (Testing Results)
- Reports files that cannot be validated

**Output:**
- Normalized assessment workbooks in output directory
- Normalization manifest (audit trail)
- Console report of normalization results

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel file reading

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl --break-system-packages

Dependencies:
    - openpyxl (Python Excel library)
    - os, sys (standard library)
    - datetime (standard library)
    - shutil (standard library)
    - argparse (standard library)
    - pathlib (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    # Normalize files in current directory
    python3 normalize_a813_assessments.py

Advanced Usage:
    # Specify source and output directories
    python3 normalize_a813_assessments.py --source ./assessments --output ./Dashboard_Sources
    
    # Automated mode (no prompts)
    python3 normalize_a813_assessments.py --source ./assessments --auto-confirm

Command-Line Options:
    --source PATH, -s PATH    Source directory containing assessment workbooks
    --output PATH, -o PATH    Output directory for normalized files
    --auto-confirm, -y        Skip confirmation prompts (for automation)

Output Files:
    Normalized assessment workbooks:
        - ISMS-IMP-A.8.13.S1.xlsx (Backup Inventory)
        - ISMS-IMP-A.8.13.S2.xlsx (Redundancy Analysis)
        - ISMS-IMP-A.8.13.S3.xlsx (RPO/RTO Compliance)
        - ISMS-IMP-A.8.13.S4.xlsx (Testing Results)
    
    Normalization manifest:
        - Normalization_Manifest_YYYYMMDD_HHMMSS.txt

Workflow Examples:

    1. Initial normalization (before first dashboard generation):
       python3 normalize_a813_assessments.py --source ./completed_assessments --output ./Dashboard_Sources
    
    2. Re-normalization after assessment updates:
       python3 normalize_a813_assessments.py --auto-confirm
    
    3. Automated integration in scripts:
       python3 normalize_a813_assessments.py --source "$ASSESSMENT_DIR" --output "$DASHBOARD_DIR" -y

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.8.13, A.8.14, A.5.30
Utility Type:         Quality Assurance - File Normalization for Dashboard Linking
Framework Version:    1.0
Script Version:       1.0
Author:               [Developer Name / Organisation]
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.13-14-5.30: BC/DR Framework Policy (Governance)
    - ISMS-IMP-A.8.13-14-5.30-S5: BC/DR Assessment Guide
    - ISMS-IMP-A.8.13.S1: Backup Inventory Assessment (Source File)
    - ISMS-IMP-A.8.13.S2: Redundancy Analysis Assessment (Source File)
    - ISMS-IMP-A.8.13.S3: RPO/RTO Compliance Matrix (Source File)
    - ISMS-IMP-A.8.13.S4: BC/DR Testing Results Tracker (Source File)
    - ISMS-IMP-A.8.13.S5: BC/DR Compliance Dashboard (Target for linking)

Related Scripts:
    - generate_a813_1_backup_inventory.py
    - generate_a813_2_redundancy_analysis.py
    - generate_a813_3_rpo_rto_compliance.py
    - generate_a813_4_testing_results.py
    - generate_a813_5_compliance_dashboard.py

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements file normalization for dashboard external workbook linking
    - Validates Document IDs in Instructions sheet
    - Creates audit manifest for traceability
    - Supports automated and interactive modes

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Why Normalization is Required:**
Excel external workbook formulas reference specific filenames:
    ='[ISMS-IMP-A.8.13.S1.xlsx]Backup_Inventory'!$B$15

If source files have date suffixes or version numbers, formula references break
when files are updated. Normalization creates stable filenames.

**Normalization vs. Renaming:**
This script COPIES files with normalized names, it does NOT rename originals:
- Original files: Unchanged, retain date/version information
- Normalized files: Copies with standardized names for dashboard linking
- Best practice: Keep both versions (original for archive, normalized for dashboard)

**Document ID Validation:**
Script validates Document ID field in Instructions sheet to ensure correct file
is being normalized. This prevents accidentally normalizing wrong files or
overwriting with incorrect assessments.

**Pre-Dashboard Requirement:**
Run this script BEFORE opening dashboard for first time:
1. Complete all four source assessments
2. Run normalization script
3. Place normalized files in same directory as dashboard
4. Open dashboard and enable external links

**Re-Running Normalization:**
Safe to re-run after assessment updates:
- Existing normalized files will be overwritten
- Manifest tracks normalization history
- No data loss in original source files

**Audit Considerations:**
Normalization manifest provides audit trail:
- Which original files were normalized
- When normalization occurred
- Document IDs validated
- Output filenames created

Keep manifests as evidence of quality assurance process.

**Data Protection:**
Normalized files are copies of source assessments:
- Inherit data classification from source files
- Handle according to organization's data classification policy
- Secure directory permissions on output folder

**Maintenance:**
Re-run normalization:
- After any updates to source assessment files
- Before dashboard refresh presentations
- As part of monthly/quarterly dashboard update process

**Quality Assurance:**
Validate normalization results:
- Check console output for any validation errors
- Verify all four expected normalized files were created
- Review manifest for any anomalies
- Test dashboard external links after normalization

**Error Handling:**
Script continues processing even if some files fail validation:
- Reports which files were successfully normalized
- Reports which files failed validation and why
- Dashboard can still function with partial normalization

**File Location Best Practices:**
Recommended directory structure:
```
BC_DR_Assessments/
├── Source/                          # Original dated/versioned files
│   ├── ISMS_Assessment_Backup_20250125.xlsx
│   ├── ISMS_Assessment_Redundancy_20250125.xlsx
│   └── ...
├── Dashboard_Sources/               # Normalized files for linking
│   ├── ISMS-IMP-A.8.13.S1.xlsx
│   ├── ISMS-IMP-A.8.13.S2.xlsx
│   └── ...
└── Dashboard/
    └── ISMS_Assessment_BCDR_Dashboard.xlsx
```

**Automation Integration:**
Script supports non-interactive mode (--auto-confirm) for integration with:
- Scheduled tasks / cron jobs
- CI/CD pipelines
- Assessment workflow automation
- Batch processing scripts

**Troubleshooting Common Issues:**

**Issue: "Document ID not found in file"**
Solution: Verify Instructions sheet exists and contains "Document ID:" field

**Issue: "Unexpected Document ID: XXX"**
Solution: File is not a recognized BC/DR assessment, check Document ID value

**Issue: "No Excel files found"**
Solution: Verify source directory path and that .xlsx files exist

**Issue: "Permission denied writing to output directory"**
Solution: Check output directory permissions and disk space

**Alternative Approaches:**
If normalization doesn't fit your workflow:
- Option 1: Manual renaming (error-prone, not recommended)
- Option 2: Modified dashboard with dynamic linking (complex Excel formulas)
- Option 3: Consolidation script instead of external linking

This normalization approach is simplest and most maintainable.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
from datetime import datetime
from pathlib import Path
import argparse
import os
import shutil
import sys

# =============================================================================
# Third-Party Imports
# =============================================================================
try:
    import openpyxl
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("ℹ️  Install with: pip install openpyxl --break-system-packages")
    sys.exit(1)


# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# CONFIGURATION
# =============================================================================

# Expected document IDs and titles from assessment workbooks
EXPECTED_DOCS = {
    "ISMS-IMP-A.8.13.S1": {
        "title": "Backup Inventory & Coverage Assessment",
        "normalized": "ISMS-IMP-A.8.13.S1.xlsx"
    },
    "ISMS-IMP-A.8.13.S2": {
        "title": "Redundancy Analysis & SPOF Assessment",
        "normalized": "ISMS-IMP-A.8.13.S2.xlsx"
    },
    "ISMS-IMP-A.8.13.S3": {
        "title": "RPO/RTO Compliance Matrix",
        "normalized": "ISMS-IMP-A.8.13.S3.xlsx"
    },
    "ISMS-IMP-A.8.13.S4": {
        "title": "BC/DR Testing Results Tracker",
        "normalized": "ISMS-IMP-A.8.13.S4.xlsx"
    },
}


# =============================================================================
# VALIDATION FUNCTIONS
# =============================================================================

def validate_workbook(filepath):
    """
    Validate workbook contains expected document ID in Instructions sheet.
    
    Args:
        filepath: Path to Excel workbook
    
    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        # Open workbook in read-only mode
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # Check for Instructions sheet
        sheet_name = None
        if "Instructions" in wb.sheetnames:
            sheet_name = "Instructions"
        elif "Instructions & Legend" in wb.sheetnames:
            sheet_name = "Instructions & Legend"
        else:
            # Try first sheet
            sheet_name = wb.sheetnames[0] if wb.sheetnames else None
        
        if not sheet_name:
            wb.close()
            return (None, None)
        
        ws = wb[sheet_name]
        
        # Look for Document ID in rows 3-25 (standardized location)
        # Format: "Document ID:" label in column A, value in column B
        for row in range(3, 25):
            cell_label = ws.cell(row=row, column=1).value
            
            if cell_label and "Document ID" in str(cell_label):
                doc_id_value = ws.cell(row=row, column=2).value
                
                if doc_id_value:
                    doc_id = str(doc_id_value).strip()
                    
                    # Check if it's one of our expected documents
                    if doc_id in EXPECTED_DOCS:
                        title = EXPECTED_DOCS[doc_id]["title"]
                        wb.close()
                        return (doc_id, title)
        
        wb.close()
        return (None, None)
        
    except Exception as e:
        print(f"    ⚠️  Error reading file: {e}")
        return (None, None)


def get_file_info(filepath):
    """
    Get file metadata.
    
    Args:
        filepath: Path object
        
    Returns:
        dict: File metadata
    """
    stat = filepath.stat()
    return {
        'size': stat.st_size,
        'modified': datetime.fromtimestamp(stat.st_mtime),
        'name': filepath.name
    }


# =============================================================================
# NORMALIZATION FUNCTIONS
# =============================================================================

def scan_directory(source_dir):
    """
    Scan directory for Excel assessment files.
    
    Args:
        source_dir: Path to source directory
        
    Returns:
        dict: Validated workbooks {doc_id: filepath}
    """
    source_path = Path(source_dir)
    
    if not source_path.exists():
        print(f"❌ Error: Source directory does not exist: {source_dir}")
        return {}
    
    # Find all Excel files
    xlsx_files = list(source_path.glob("*.xlsx"))
    
    if not xlsx_files:
        print(f"⚠️  No Excel files found in {source_dir}")
        return {}
    
    print(f"🔍 Scanning {len(xlsx_files)} Excel file(s) in {source_dir}...")
    print()
    
    validated = {}
    
    for filepath in xlsx_files:
        # Skip temporary Excel files
        if filepath.name.startswith('~$'):
            continue
            
        print(f"  Checking: {filepath.name}")
        
        doc_id, title = validate_workbook(filepath)
        
        if doc_id:
            print(f"    ✅ Valid: {doc_id} - {title}")
            
            # Check for duplicates
            if doc_id in validated:
                print(f"    ⚠️  WARNING: Duplicate found (keeping first occurrence)")
            else:
                validated[doc_id] = filepath
        else:
            print(f"    ⏭️  Skipped (no valid Document ID found)")
        
        print()
    
    return validated


def normalize_files(validated_files, output_dir, auto_confirm=False):
    """
    Copy validated files to normalized names.
    
    Args:
        validated_files: Dict of {doc_id: filepath}
        output_dir: Output directory path
        auto_confirm: Skip confirmation prompts
        
    Returns:
        dict: Normalization results
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    results = {
        'success': [],
        'failed': [],
        'skipped': []
    }
    
    print("=" * 80)
    print("NORMALIZING FILES")
    print("=" * 80)
    print()
    
    for doc_id, source_file in validated_files.items():
        normalized_name = EXPECTED_DOCS[doc_id]["normalized"]
        dest_file = output_path / normalized_name
        
        print(f"📄 {doc_id}")
        print(f"   Source: {source_file.name}")
        print(f"   Dest:   {normalized_name}")
        
        try:
            # Check if destination exists
            if dest_file.exists():
                if not auto_confirm:
                    response = input(f"   ⚠️  {normalized_name} exists. Overwrite? [y/N]: ")
                    if response.lower() != 'y':
                        print(f"   ⏭️  Skipped")
                        results['skipped'].append(doc_id)
                        print()
                        continue
            
            # Copy file
            shutil.copy2(source_file, dest_file)
            print(f"   ✅ Copied successfully")
            results['success'].append({
                'doc_id': doc_id,
                'source': source_file.name,
                'dest': normalized_name
            })
            
        except Exception as e:
            print(f"   ❌ Error: {e}")
            results['failed'].append(doc_id)
        
        print()
    
    return results


def create_manifest(results, output_dir):
    """
    Create normalization audit manifest.
    
    Args:
        results: Normalization results dict
        output_dir: Output directory path
    """
    manifest_path = Path(output_dir) / "normalization_manifest.txt"
    
    with open(manifest_path, 'w') as f:
        f.write("=" * 80 + "\n")
        f.write("ISMS ASSESSMENT NORMALIZATION MANIFEST\n")
        f.write("Control A.8.13 - Information Backup\n")
        f.write("=" * 80 + "\n\n")
        
        f.write(f"Normalization Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Output Directory: {output_dir}\n\n")
        
        f.write("SUCCESSFUL NORMALIZATIONS:\n")
        f.write("-" * 80 + "\n")
        for item in results['success']:
            f.write(f"Document ID: {item['doc_id']}\n")
            f.write(f"  Source: {item['source']}\n")
            f.write(f"  Normalized: {item['dest']}\n\n")
        
        if results['skipped']:
            f.write("\nSKIPPED FILES:\n")
            f.write("-" * 80 + "\n")
            for doc_id in results['skipped']:
                f.write(f"  {doc_id}\n")
        
        if results['failed']:
            f.write("\nFAILED NORMALIZATIONS:\n")
            f.write("-" * 80 + "\n")
            for doc_id in results['failed']:
                f.write(f"  {doc_id}\n")
        
        f.write("\n" + "=" * 80 + "\n")
    
    print(f"📝 Creating normalization manifest...")
    print(f"   ✅ Manifest created: {manifest_path.name}")
    print()


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main execution function"""
    
    parser = argparse.ArgumentParser(
        description='Normalize A.8.13 assessment files for dashboard linking',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Normalize files in current directory
  python3 normalize_a813_assessments.py
  
  # Specify source and output directories
  python3 normalize_a813_assessments.py --source ./assessments --output ./Dashboard_Sources
  
  # Automated mode (no prompts)
  python3 normalize_a813_assessments.py --source ./assessments --auto-confirm
        """
    )
    
    parser.add_argument(
        '--source', '-s',
        help='Source directory containing assessment workbooks (default: current directory)'
    )
    
    parser.add_argument(
        '--output', '-o',
        help='Output directory for normalized files (default: ./Dashboard_Sources)'
    )
    
    parser.add_argument(
        '--auto-confirm', '-y',
        action='store_true',
        help='Skip confirmation prompts (for automation)'
    )
    
    args = parser.parse_args()
    
    # Set defaults
    source_dir = args.source if args.source else os.getcwd()
    output_dir = args.output if args.output else os.path.join(os.getcwd(), 'Dashboard_Sources')
    
    # Print header
    print("=" * 80)
    print("ISMS ASSESSMENT FILE NORMALIZATION UTILITY")
    print("ISO/IEC 27001:2022 - Control A.8.13: Information Backup")
    print("=" * 80)
    print()
    print("This script prepares assessment workbooks for dashboard linking by:")
    print("  1. Scanning for completed assessment files")
    print("  2. Validating document IDs in Instructions sheet")
    print("  3. Copying to normalized filenames (no dates/versions)")
    print("  4. Creating audit manifest for traceability")
    print()
    print()
    print(f"📂 Source directory: {source_dir}")
    print(f"📂 Output directory: {output_dir}")
    print()
    
    # Scan and validate
    validated = scan_directory(source_dir)
    
    if not validated:
        print("❌ No valid assessment files found")
        sys.exit(1)
    
    # Show summary
    print("=" * 80)
    print("NORMALIZATION SUMMARY")
    print("=" * 80)
    print()
    print(f"Found {len(validated)} of {len(EXPECTED_DOCS)} required assessment workbooks:")
    print()
    
    for doc_id in sorted(EXPECTED_DOCS.keys()):
        if doc_id in validated:
            print(f"  ✅ {doc_id}")
            print(f"     Title:      {EXPECTED_DOCS[doc_id]['title']}")
            print(f"     Source:     {validated[doc_id].name}")
            print(f"     Normalized: {EXPECTED_DOCS[doc_id]['normalized']}")
        else:
            print(f"  ❌ {doc_id}")
            print(f"     Title:      {EXPECTED_DOCS[doc_id]['title']}")
            print(f"     Status:     NOT FOUND")
        print()
    
    print(f"Output directory: {output_dir}")
    print()
    
    # Confirm before proceeding
    if not args.auto_confirm:
        response = input("Proceed with normalization? [y/N]: ")
        if response.lower() != 'y':
            print("❌ Normalization cancelled")
            sys.exit(0)
        print()
    
    # Normalize files
    results = normalize_files(validated, output_dir, args.auto_confirm)
    
    # Create manifest
    create_manifest(results, output_dir)
    
    # Final summary
    print("=" * 80)
    print("NORMALIZATION COMPLETE")
    print("=" * 80)
    print()
    print(f"✅ Successfully normalized {len(results['success'])}/{len(validated)} workbooks")
    print(f"📂 Output directory: {output_dir}")
    print()
    
    if len(results['success']) == len(EXPECTED_DOCS):
        print(f"✅ All {len(EXPECTED_DOCS)} required assessments normalized")
        print("   Ready for dashboard generation")
    else:
        missing = len(EXPECTED_DOCS) - len(results['success'])
        print(f"⚠️  {missing} assessment(s) still missing")
    
    print()
    print("NEXT STEPS:")
    print(f"  1. Generate dashboard: python3 generate_a813_5_dashboard.py")
    print(f"  2. Place dashboard in: {output_dir}")
    print("  3. Open dashboard and click 'Update Links' when prompted")
    print()
    print("=" * 80)
    print()


if __name__ == '__main__':
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION (syntax validated, structure verified)
# QA_TOOL: Claude Code Standardization
# =============================================================================
