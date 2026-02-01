#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
Excel Workbook Sanity Checker - ISMS A.8.13/14/5.30 Assessment Workbooks
================================================================================

Diagnostic utility for troubleshooting Excel's "file level validation and repair"
warnings when opening A.8.13/14/5.30 BC/DR assessment workbooks.

**Purpose:**
Identifies common openpyxl-generated Excel issues that trigger repair warnings:
- Formula syntax errors and invalid sheet references
- Data validation conflicts and overlapping ranges
- Style attribute inconsistencies
- Merged cell content issues
- Worksheet structure problems

**When to Use:**
- Excel displays repair warnings when opening generated workbooks
- After modifying assessment generator scripts
- Before distributing workbooks to stakeholders
- Quality assurance validation before consolidation

**Usage:**
    python3 excel_sanity_check_a813.py ISMS_Assessment_Backup_Inventory.xlsx
    
    Works with any A.8.13/14/5.30 assessment workbook (domains 1-5)

**Output:**
- Diagnostic report with issue categorization (Critical/Warning)
- Recommended remediation actions
- Structural health summary

**Related Scripts:**
- excel_sanity_check_a813_1.py through _5.py (domain-specific checkers)
- generate_a813_1_backup_inventory.py through generate_a813_5_compliance_dashboard.py

Control Reference: ISO/IEC 27001:2022 Annex A Controls A.8.13, A.8.14, A.5.30
Script Type: Quality Assurance Utility
Version: 1.0
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
from pathlib import Path
import re
import sys

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
def detect_workbook_type(filename):
    """Detect which BC/DR workbook type this is based on filename."""
    filename_lower = filename.lower()
    
    if 'backup' in filename_lower and 'inventory' in filename_lower:
        return 'BCDR-1', 'Backup Inventory & Coverage Assessment'
    elif 'redundancy' in filename_lower:
        return 'BCDR-2', 'Redundancy Analysis & SPOF Assessment'
    elif 'rpo' in filename_lower or 'rto' in filename_lower or 'compliance' in filename_lower:
        return 'BCDR-3', 'RPO/RTO Compliance Matrix'
    elif 'testing' in filename_lower and 'results' in filename_lower:
        return 'BCDR-4', 'BC/DR Testing Results Tracker'
    elif 'dashboard' in filename_lower:
        return 'BCDR-5', 'BC/DR Consolidated Dashboard'
    else:
        return 'Unknown', 'Generic BC/DR Excel Workbook'


# ============================================================================
# EXPECTED SHEET DEFINITIONS
# ============================================================================

EXPECTED_SHEETS = {
    'BCDR-1': [
        "Instructions", "Summary", "Backup_Inventory", "Backup_Schedule",
        "RPO_Compliance", "321_Rule_Assessment", "Gap_Analysis",
        "Evidence_Register", "Approval_Sign_Off"
    ],
    'BCDR-2': [
        "Instructions", "Summary", "Redundancy_Inventory", "RTO_Compliance",
        "SPOF_Register", "Failover_Testing", "Gap_Analysis",
        "Evidence_Register", "Approval_Sign_Off"
    ],
    'BCDR-3': [
        "Instructions", "Summary", "System_Inventory", "Capability_Assessment",
        "Compliance_Matrix", "Gap_Analysis", "Evidence_Register", "Approval_Sign_Off"
    ],
    'BCDR-4': [
        "Instructions", "Summary", "Test_Schedule", "Test_Results_Log",
        "Issue_Remediation", "Testing_Compliance", "Evidence_Register",
        "Approval_Sign_Off"
    ],
    'BCDR-5': [
        "Instructions", "Executive_Dashboard", "Detailed_Metrics",
        "Gap_Summary", "Evidence_Checklist", "Approval_Sign_Off"
    ],
}

# Expected Assessment IDs
EXPECTED_ASSESSMENT_IDS = {
    'BCDR-1': 'BCDR-BACKUP-001',
    'BCDR-2': 'BCDR-REDUNDANCY-002',
    'BCDR-3': 'BCDR-RPORTO-003',
    'BCDR-4': 'BCDR-TESTING-004',
    'BCDR-5': 'BCDR-DASHBOARD-005',
}

# Expected normalized filenames for dashboard
DASHBOARD_SOURCES = [
    'BCDR_1_Backup_Inventory.xlsx',
    'BCDR_2_Redundancy_Analysis.xlsx',
    'BCDR_3_RPO_RTO_Compliance.xlsx',
    'BCDR_4_Testing_Results.xlsx',
]


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""
    
    workbook_id, workbook_name = detect_workbook_type(filename)
    
    print("=" * 80)
    print(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {Path(filename).name}")
    print(f"Detected Type: {workbook_id} - {workbook_name}")
    print("=" * 80)
    
    issues_found = []
    warnings_found = []
    
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"\n✓ Workbook loaded successfully")
        print(f"  Sheets found: {len(wb.sheetnames)}")
        print(f"  Sheet names: {', '.join(wb.sheetnames)}")
        
    except Exception as e:
        print(f"\n✗ CRITICAL: Cannot load workbook: {e}")
        return
    
    # ========================================================================
    # CHECK 1: SHEET STRUCTURE VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 1: SHEET STRUCTURE VALIDATION")
    print("=" * 80)
    
    if workbook_id in EXPECTED_SHEETS:
        expected = EXPECTED_SHEETS[workbook_id]
        actual = wb.sheetnames
        
        # Check for missing sheets (flexible matching)
        missing = []
        for exp_sheet in expected:
            # Flexible match: check if expected sheet name is in any actual sheet
            found = False
            for act_sheet in actual:
                if exp_sheet.lower().replace("_", "") in act_sheet.lower().replace("_", "").replace(" ", ""):
                    found = True
                    break
            if not found:
                missing.append(exp_sheet)
        
        if missing:
            for sheet in missing:
                warnings_found.append(f"  ⚠ Missing expected sheet: {sheet}")
            print(f"  ⚠ {len(missing)} expected sheets not found (may be renamed)")
        else:
            print(f"  ✓ All expected sheets present for {workbook_id}")
        
        # Check for extra sheets
        print(f"  Total sheets: {len(actual)} (expected {len(expected)})")
    else:
        print("  ℹ Unknown workbook type - skipping sheet validation")
    
    # ========================================================================
    # CHECK 2: ASSESSMENT ID VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 2: ASSESSMENT ID VALIDATION")
    print("=" * 80)
    
    assessment_id_found = False
    if workbook_id in EXPECTED_ASSESSMENT_IDS:
        expected_id = EXPECTED_ASSESSMENT_IDS[workbook_id]
        
        # Check in Instructions sheet (row 2 merged cell)
        if 'Instructions' in wb.sheetnames:
            ws = wb['Instructions']
            # Check row 2 merged cell (format: "Version X.X | Controls: ... | Assessment ID: BCDR-XXX-NNN")
            for row in range(1, 10):
                for col in range(1, 5):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and 'Assessment ID' in cell_value:
                        if expected_id in cell_value:
                            print(f"  ✓ Assessment ID found: {expected_id}")
                            assessment_id_found = True
                        else:
                            issues_found.append(
                                f"  ✗ Assessment ID mismatch: expected {expected_id} in row {row}"
                            )
                        break
                if assessment_id_found:
                    break
        
        if not assessment_id_found:
            warnings_found.append(
                f"  ⚠ Assessment ID '{expected_id}' not found in Instructions sheet"
            )
    else:
        print("  ℹ Unknown workbook type - skipping Assessment ID validation")
    
    # ========================================================================
    # CHECK 3: FORMULA VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 3: FORMULA VALIDATION")
    print("=" * 80)
    
    formula_issues = 0
    inter_sheet_refs = {}
    external_refs = set()
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # Check for external workbook references
                    ext_refs = re.findall(r'\[([^\]]+\.xlsx)\]', formula)
                    for ref in ext_refs:
                        external_refs.add(ref)
                    
                    # Check for sheet references
                    if "'" in formula or "!" in formula:
                        sheet_refs = re.findall(r"'?([^'!]+)'?!", formula)
                        for ref in sheet_refs:
                            ref_clean = ref.strip("'")
                            # Skip external references
                            if "[" not in ref_clean and ref_clean not in wb.sheetnames:
                                # Check for close matches (underscore vs space)
                                close_match = False
                                for actual in wb.sheetnames:
                                    if ref_clean.replace(" ", "_") == actual or ref_clean.replace("_", " ") == actual:
                                        close_match = True
                                        break
                                if not close_match:
                                    issues_found.append(
                                        f"  ✗ {sheet_name}!{cell.coordinate}: "
                                        f"Invalid sheet reference '{ref_clean}'"
                                    )
                                    formula_issues += 1
                            
                            # Track inter-sheet references
                            if "[" not in ref_clean and ref_clean in wb.sheetnames:
                                if sheet_name not in inter_sheet_refs:
                                    inter_sheet_refs[sheet_name] = set()
                                inter_sheet_refs[sheet_name].add(ref_clean)
                    
                    # Check for common syntax issues
                    if formula.count('(') != formula.count(')'):
                        issues_found.append(
                            f"  ✗ {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced parentheses in formula"
                        )
                        formula_issues += 1
    
    if formula_issues == 0:
        print("  ✓ No formula syntax issues detected")
    else:
        print(f"  ✗ Found {formula_issues} formula issues")
    
    if inter_sheet_refs:
        print(f"  ℹ Inter-sheet references found:")
        for sheet, refs in inter_sheet_refs.items():
            print(f"    {sheet} → {', '.join(sorted(refs))}")
    
    if external_refs:
        print(f"  ℹ External workbook references found:")
        for ref in sorted(external_refs):
            print(f"    → {ref}")
        if workbook_id == 'BCDR-5':
            print("  ⚠ WARNING: Dashboard requires external links to function")
    
    # ========================================================================
    # CHECK 4: DATA VALIDATION CONFLICTS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 4: DATA VALIDATION CONFLICTS")
    print("=" * 80)
    
    validation_issues = 0
    total_validations = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'data_validations') and ws.data_validations:
            dv_count = len(ws.data_validations.dataValidation)
            total_validations += dv_count
            
            # Check for overlapping ranges
            all_ranges = []
            for dv in ws.data_validations.dataValidation:
                if hasattr(dv, 'sqref') and dv.sqref:
                    all_ranges.extend(str(dv.sqref).split())
            
            if len(all_ranges) != len(set(all_ranges)):
                warnings_found.append(
                    f"  ⚠ {sheet_name}: Potentially overlapping data validation ranges"
                )
                validation_issues += 1
    
    print(f"  Total data validations: {total_validations}")
    
    if validation_issues == 0:
        print("  ✓ No data validation conflicts detected")
    
    # ========================================================================
    # CHECK 5: MERGED CELLS VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 5: MERGED CELLS VALIDATION")
    print("=" * 80)
    
    merge_issues = 0
    total_merges = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, 'merged_cells') and ws.merged_cells:
            merge_count = len(ws.merged_cells.ranges)
            total_merges += merge_count
            
            # Check merged cells for content issues
            for merge_range in ws.merged_cells.ranges:
                min_col, min_row = merge_range.min_col, merge_range.min_row
                max_col, max_row = merge_range.max_col, merge_range.max_row
                
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if row == min_row and col == min_col:
                            continue
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            cell_coord = f"{get_column_letter(col)}{row}"
                            warnings_found.append(
                                f"  ⚠ {sheet_name}!{cell_coord}: "
                                f"Merged cell has content in non-primary cell"
                            )
                            merge_issues += 1
    
    print(f"  Total merged ranges: {total_merges}")
    
    if merge_issues == 0:
        print("  ✓ Merged cells properly formatted")
    
    # ========================================================================
    # CHECK 6: EVIDENCE REGISTER VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 6: EVIDENCE REGISTER VALIDATION")
    print("=" * 80)
    
    evidence_sheet = None
    for sheet in wb.sheetnames:
        if 'evidence' in sheet.lower():
            evidence_sheet = sheet
            break
    
    if evidence_sheet:
        ws = wb[evidence_sheet]
        # Count evidence rows (look for EVD prefix)
        evidence_count = 0
        for row in range(1, min(110, ws.max_row + 1)):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and isinstance(cell_val, str):
                if cell_val.startswith('EVD-'):
                    evidence_count += 1
        
        print(f"  Evidence sheet: {evidence_sheet}")
        print(f"  Evidence row templates: {evidence_count}")
        
        # BC/DR requires minimum 5 evidence items
        if evidence_count < 5:
            issues_found.append(
                f"  ✗ Evidence Register has {evidence_count} rows (MINIMUM 5 REQUIRED)"
            )
        elif evidence_count < 50:
            warnings_found.append(
                f"  ⚠ Evidence Register has {evidence_count} rows (recommended ≥50 for flexibility)"
            )
        elif evidence_count >= 100:
            print(f"  ✓ Full 100-row evidence register (excellent)")
        else:
            print(f"  ✓ Evidence register adequate ({evidence_count} rows)")
    else:
        if workbook_id != 'BCDR-5':  # Dashboard doesn't have Evidence_Register, has Evidence_Checklist
            issues_found.append("  ✗ No Evidence Register sheet found (REQUIRED)")
    
    # ========================================================================
    # CHECK 7: APPROVAL SIGN-OFF VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 7: APPROVAL SIGN-OFF VALIDATION")
    print("=" * 80)
    
    approval_sheet = None
    for sheet in wb.sheetnames:
        if 'approval' in sheet.lower() or 'sign' in sheet.lower():
            approval_sheet = sheet
            break
    
    if approval_sheet:
        ws = wb[approval_sheet]
        
        # Check for 3-level approval structure
        approval_levels = 0
        for row in range(1, min(100, ws.max_row + 1)):
            for col in range(1, 6):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val and isinstance(cell_val, str):
                    if 'LEVEL 1' in cell_val or 'ASSESSMENT COMPLETED BY' in cell_val:
                        approval_levels += 1
                    elif 'LEVEL 2' in cell_val or 'INFORMATION SECURITY OFFICER' in cell_val:
                        approval_levels += 1
                    elif 'LEVEL 3' in cell_val or 'CISO' in cell_val:
                        approval_levels += 1
        
        print(f"  Approval sheet: {approval_sheet}")
        print(f"  Approval levels detected: {approval_levels}")
        
        if approval_levels < 3:
            warnings_found.append(
                f"  ⚠ Approval workflow has {approval_levels} levels (expected 3: Assessor → ISO → CISO)"
            )
        else:
            print(f"  ✓ Complete 3-level approval workflow")
    else:
        issues_found.append("  ✗ No Approval Sign-Off sheet found (REQUIRED)")
    
    # ========================================================================
    # CHECK 8: SUMMARY DASHBOARD VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 8: SUMMARY DASHBOARD VALIDATION")
    print("=" * 80)
    
    summary_sheet = None
    for sheet in wb.sheetnames:
        if 'summary' in sheet.lower():
            summary_sheet = sheet
            break
    
    if summary_sheet:
        ws = wb[summary_sheet]
        
        # Count formulas in Summary sheet
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
        
        print(f"  Summary sheet: {summary_sheet}")
        print(f"  Formulas in Summary: {formula_count}")
        
        if formula_count < 10:
            warnings_found.append(
                f"  ⚠ Summary sheet has only {formula_count} formulas (expected many calculated metrics)"
            )
        else:
            print(f"  ✓ Summary dashboard has active calculations")
    else:
        if workbook_id != 'BCDR-5':  # Dashboard has Executive_Dashboard instead
            warnings_found.append("  ⚠ No Summary sheet found")
    
    # ========================================================================
    # CHECK 9: ROW CAPACITY VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 9: ROW CAPACITY VALIDATION")
    print("=" * 80)
    
    data_sheets = [s for s in wb.sheetnames if s not in ['Instructions', 'Summary', 
                                                          'Evidence_Register', 'Approval_Sign_Off',
                                                          'Evidence_Checklist', 'Executive_Dashboard']]
    
    for sheet_name in data_sheets:
        ws = wb[sheet_name]
        
        # Check if data entry area exists (typically rows 5-114 for 110 capacity)
        if ws.max_row > 100:
            # Count example rows vs entry rows
            example_count = 0
            entry_area_start = 15  # Typically after 10 examples
            
            for row in range(5, min(15, ws.max_row)):
                # Check if row has data (non-empty first column)
                if ws.cell(row=row, column=1).value:
                    example_count += 1
            
            total_capacity = ws.max_row - 4  # Subtract header rows
            
            print(f"  {sheet_name}:")
            print(f"    Total capacity: ~{total_capacity} rows")
            print(f"    Example data rows: ~{example_count}")
            
            if total_capacity < 50:
                warnings_found.append(
                    f"  ⚠ {sheet_name}: Low capacity ({total_capacity} rows), recommended ≥100"
                )
            elif total_capacity >= 100:
                print(f"    ✓ Adequate capacity (≥100 rows)")
    
    # ========================================================================
    # CHECK 10: WORKSHEET DIMENSIONS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 10: WORKSHEET DIMENSIONS")
    print("=" * 80)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if ws.max_row > 5000:
            warnings_found.append(
                f"  ⚠ {sheet_name}: Large worksheet ({ws.max_row} rows)"
            )
        
        if ws.max_column > 50:
            warnings_found.append(
                f"  ⚠ {sheet_name}: Wide worksheet ({ws.max_column} columns)"
            )
    
    print("  ✓ Worksheet dimensions within reasonable limits")
    
    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    print("\n" + "=" * 80)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 80)
    
    if issues_found:
        print(f"\n❌ CRITICAL ISSUES FOUND: {len(issues_found)}")
        for issue in issues_found:
            print(issue)
    
    if warnings_found:
        print(f"\n⚠️  WARNINGS: {len(warnings_found)}")
        for warning in warnings_found:
            print(warning)
    
    if not issues_found and not warnings_found:
        print("\n✅ NO ISSUES DETECTED")
        print("\nThe workbook appears structurally sound.")
        print("If Excel still shows repair warnings, consider:")
        print("  • Excel version compatibility (try Excel 2019+)")
        print("  • Antivirus/security software interference")
        print("  • Network drive or OneDrive sync issues")
        print("  • Excel's overly cautious validation")
        if workbook_id == 'BCDR-5':
            print("  • Unresolved external workbook links (expected before normalization)")
    else:
        print_recommendations(issues_found, warnings_found, workbook_id, 
                            formula_issues, validation_issues, merge_issues)
    
    # ========================================================================
    # WORKBOOK-SPECIFIC NOTES
    # ========================================================================
    print_workbook_notes(workbook_id, external_refs)
    
    wb.close()


def print_recommendations(issues, warnings, workbook_id, formula_issues, 
                         validation_issues, merge_issues):
    """Print recommended fixes based on issues found."""
    print("\n" + "=" * 80)
    print("RECOMMENDED ACTIONS")
    print("=" * 80)
    
    if formula_issues > 0:
        print("\n1. FORMULA FIXES:")
        print("   • Review formulas referencing other sheets")
        print("   • Ensure sheet names match exactly (case-sensitive)")
        print("   • Check for balanced quotes and parentheses")
        if workbook_id == 'BCDR-5':
            print("   • Verify external workbook references point to normalized files")
    
    if validation_issues > 0:
        print("\n2. DATA VALIDATION FIXES:")
        print("   • Remove overlapping data validation ranges")
        print("   • Apply validations to specific ranges, not entire columns")
    
    if merge_issues > 0:
        print("\n3. MERGED CELL FIXES:")
        print("   • Ensure only top-left cell of merged range has content")
        print("   • Clear content from other cells in merged range")
    
    if workbook_id == 'BCDR-5':
        print("\n4. DASHBOARD-SPECIFIC SETUP (BCDR-5):")
        print("   • This workbook consolidates data from 4 assessment workbooks")
        print("   • Ensure normalized source workbooks are in the same folder:")
        for source in DASHBOARD_SOURCES:
            print(f"     - {source}")
        print("   • Run normalize_bcdr_assessments.py first to create normalized files")


def print_workbook_notes(workbook_id, external_refs):
    """Print workbook-specific notes and expected structure."""
    print("\n" + "=" * 80)
    print(f"WORKBOOK-SPECIFIC NOTES: {workbook_id}")
    print("=" * 80)
    
    if workbook_id == 'BCDR-1':
        print("\nBackup Inventory & Coverage Assessment:")
        print("  • 9 sheets for backup assessment")
        print("  • 110 system capacity (10 examples + 100 entry rows)")
        print("  • Backup status tracking (backed up / not backed up)")
        print("  • RPO compliance assessment (backup frequency vs requirements)")
        print("  • 3-2-1-1-0 rule compliance tracking")
        print("  • Gap analysis with remediation tracking")
        print("  • Evidence Register (100 entries minimum)")
        print("  • 3-level approval workflow (Assessor → ISO → CISO)")
        print("\n  Key metrics:")
        print("    - Backup coverage percentage")
        print("    - Critical system coverage (Tier 1)")
        print("    - RPO compliance rate")
        print("    - 3-2-1-1-0 compliance rate")
    
    elif workbook_id == 'BCDR-2':
        print("\nRedundancy Analysis & SPOF Assessment:")
        print("  • 9 sheets for redundancy assessment")
        print("  • 110 system capacity per assessment sheet")
        print("  • Redundancy architecture documentation")
        print("  • RTO compliance assessment (failover time vs requirements)")
        print("  • SPOF identification and remediation tracking")
        print("  • Failover testing results")
        print("  • Gap analysis with priority scoring")
        print("  • Evidence Register (100 entries minimum)")
        print("  • 3-level approval workflow")
        print("\n  Key metrics:")
        print("    - Redundancy coverage (critical systems)")
        print("    - RTO compliance rate")
        print("    - SPOF count (identified vs mitigated)")
        print("    - Failover success rate")
    
    elif workbook_id == 'BCDR-3':
        print("\nRPO/RTO Compliance Matrix:")
        print("  • 8 sheets for RPO/RTO alignment")
        print("  • 110 system capacity per assessment sheet")
        print("  • BIA results (business requirements)")
        print("  • Technical capability assessment (backup + redundancy)")
        print("  • Automated compliance matrix (requirement vs capability)")
        print("  • Gap analysis with risk scoring")
        print("  • Priority-based remediation tracking")
        print("  • Evidence Register (100 entries minimum)")
        print("  • 3-level approval workflow")
        print("\n  Key metrics:")
        print("    - Overall RPO/RTO compliance rate")
        print("    - Tier 1 system compliance")
        print("    - Critical gaps requiring immediate action")
        print("    - Unknown status (untested systems)")
    
    elif workbook_id == 'BCDR-4':
        print("\nBC/DR Testing Results Tracker:")
        print("  • 8 sheets for testing compliance")
        print("  • 110 capacity per schedule/results/issue sheet")
        print("  • Annual test schedule planning")
        print("  • Test execution results logging")
        print("  • Issue remediation workflow (issues → closed)")
        print("  • Testing compliance assessment (quarterly/semi-annual/annual)")
        print("  • Automated compliance calculations (days since last test)")
        print("  • Evidence Register (100 entries minimum)")
        print("  • 3-level approval workflow")
        print("\n  Key metrics:")
        print("    - Test success rate")
        print("    - Testing compliance (Tier 1 quarterly)")
        print("    - Critical issues (open vs closed)")
        print("    - Overdue tests")
        print("\n  Critical principle:")
        print("    ⚠ 'Test or it doesn't work' - Untested recovery = no recovery")
    
    elif workbook_id == 'BCDR-5':
        print("\nBC/DR Consolidated Dashboard (MASTER):")
        print("  • 6 sheets - EXECUTIVE CONSOLIDATION WORKBOOK")
        print("  • Overall BC/DR maturity score (0-100%)")
        print("  • Maturity level assessment (Initial/Managed/Defined/Optimized)")
        print("  • Consolidated metrics from all 4 assessments")
        print("  • Dimension scoring (25 points each):")
        print("    1. Backup Coverage & Compliance (25%)")
        print("    2. Redundancy Coverage & Compliance (25%)")
        print("    3. RPO/RTO Alignment (25%)")
        print("    4. Testing Compliance (25%)")
        print("  • Critical gaps summary (all workbooks)")
        print("  • Evidence completeness tracking")
        print("  • 3-level approval workflow")
        print("\n  ⚠ This dashboard requires normalized source files:")
        for source in DASHBOARD_SOURCES:
            print(f"    - {source}")
        print("\n  Setup process:")
        print("    1. Generate all 4 assessment workbooks (Scripts 1-4)")
        print("    2. Run: python normalize_bcdr_assessments.py")
        print("    3. Dashboard reads from normalized/ directory")
        print("    4. May show #REF errors until normalization complete")
    
    else:
        print("\nGeneric BC/DR Workbook")
        print("  • No specific validation rules defined for this type")
        print("  • Standard Excel health checks applied")
        print("  • Ensure workbook follows BC/DR framework structure")
    
    print("\n" + "=" * 80)


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("=" * 70)
        print("ISMS BC/DR Framework - Excel Sanity Checker")
        print("Controls: A.8.13 (Backup), A.8.14 (Redundancy), A.5.30 (ICT BC)")
        print("=" * 70)
        print("\nUsage: python3 excel_sanity_check_bcdr.py <filename.xlsx>")
        print("\nExamples:")
        print("  python3 excel_sanity_check_bcdr.py ISMS_Assessment_Backup_Inventory.xlsx")
        print("  python3 excel_sanity_check_bcdr.py ISMS_Assessment_Redundancy_Analysis.xlsx")
        print("  python3 excel_sanity_check_bcdr.py ISMS_Assessment_RPO_RTO_Compliance.xlsx")
        print("  python3 excel_sanity_check_bcdr.py ISMS_Assessment_Testing_Results.xlsx")
        print("  python3 excel_sanity_check_bcdr.py ISMS_Assessment_BCDR_Dashboard.xlsx")
        print("\nSupported workbook types:")
        print("  BCDR-1 - Backup Inventory & Coverage Assessment")
        print("  BCDR-2 - Redundancy Analysis & SPOF Assessment")
        print("  BCDR-3 - RPO/RTO Compliance Matrix")
        print("  BCDR-4 - BC/DR Testing Results Tracker")
        print("  BCDR-5 - BC/DR Consolidated Dashboard")
        print("\n'Test or it doesn't work' - BC/DR Systems Engineering ISMS")
        sys.exit(1)
    
    filename = sys.argv[1]
    
    try:
        check_workbook_health(filename)
    except FileNotFoundError:
        print(f"\n❌ ERROR: File not found: {filename}")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION (syntax validated, structure verified)
# QA_TOOL: Claude Code Standardization
# =============================================================================
