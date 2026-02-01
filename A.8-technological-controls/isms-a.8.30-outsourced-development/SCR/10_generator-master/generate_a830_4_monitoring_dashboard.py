#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.30.4 - Monitoring and Exceptions Dashboard
================================================================================

ISO/IEC 27001:2022 Control A.8.30: Outsourced Development
Assessment Domain 4 of 4: Monitoring and Exceptions Dashboard

This script generates a comprehensive Excel dashboard workbook that consolidates
outsourced development security status, tracks exceptions, and provides executive
reporting on vendor security compliance.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.30.4"
WORKBOOK_NAME = "Monitoring and Exceptions Dashboard"
CONTROL_ID = "A.8.30"
CONTROL_NAME = "Outsourced Development"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions sheet."""
    ws.title = "Instructions"

    instructions = [
        ["ISMS-IMP-A.8.30.4 - Monitoring and Exceptions Dashboard"],
        [""],
        ["PURPOSE"],
        ["This dashboard consolidates outsourced development security status,"],
        ["tracks exceptions, and provides executive reporting per ISMS-POL-A.8.30."],
        [""],
        ["SHEETS"],
        ["1. Instructions - This guidance sheet"],
        ["2. Executive Dashboard - Key metrics and compliance status"],
        ["3. Vendor Performance - Performance scores by vendor"],
        ["4. Exception Register - Security exception tracking"],
        ["5. Monitoring Log - Ongoing monitoring activity records"],
        ["6. Incident Log - Security incident tracking"],
        ["7. Compliance Score - Score calculation breakdown"],
        [""],
        ["REPORTING CADENCE"],
        ["- Executive Dashboard: Monthly"],
        ["- Vendor Performance: Quarterly"],
        ["- Exception Status: Monthly"],
        ["- Compliance Score Trend: Quarterly"],
        ["- Incident Summary: As needed + Quarterly"],
        [""],
        ["COMPLIANCE SCORE INTERPRETATION"],
        ["- >= 90%: Compliant (Green)"],
        ["- 70-89%: Needs Improvement (Yellow)"],
        ["- < 70%: Non-Compliant (Red) - Escalation required"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SHEETS", "REPORTING CADENCE", "COMPLIANCE SCORE INTERPRETATION"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 80


def create_executive_dashboard_sheet(ws):
    """Create the Executive Dashboard sheet."""
    ws.title = "Executive Dashboard"

    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="Outsourced Development Security Dashboard")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")

    # Report date
    ws.cell(row=2, column=1, value=f"Report Date: {GENERATED_DATE}")
    ws.cell(row=2, column=1).font = Font(italic=True)

    # Metrics headers
    headers = ["Metric", "Description", "Target", "Current", "Trend"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Dashboard metrics
    metrics = [
        ("Approved Vendors", "Vendors in approved registry", "N/A", "[Count]"),
        ("Active Contracts", "Contracts currently active", "N/A", "[Count]"),
        ("Pending Assessments", "Vendors awaiting assessment", "0", "[Count]"),
        ("Overdue Reassessments", "Annual reassessments overdue", "0", "[Count]"),
        ("Contract Clause Compliance", "Contracts with all security clauses", "100%", "[%]"),
        ("SLA Compliance (Critical)", "Critical vulns fixed within SLA", ">=95%", "[%]"),
        ("SLA Compliance (High)", "High vulns fixed within SLA", ">=90%", "[%]"),
        ("Security Testing Completion", "Deliverables with complete testing", "100%", "[%]"),
        ("SBOM Compliance", "Deliverables with SBOM received", "100%", "[%]"),
        ("Active Exceptions", "Open security exceptions", "[Track]", "[Count]"),
        ("Overdue Exceptions", "Exceptions past expiry", "0", "[Count]"),
        ("Overall Compliance Score", "Weighted compliance score", ">=90%", "[%]"),
    ]

    row = 5
    for metric, desc, target, current in metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=2, value=desc).border = THIN_BORDER
        ws.cell(row=row, column=3, value=target).border = THIN_BORDER
        cell_current = ws.cell(row=row, column=4, value=current)
        cell_current.fill = INPUT_FILL
        cell_current.border = THIN_BORDER
        cell_trend = ws.cell(row=row, column=5)
        cell_trend.fill = INPUT_FILL
        cell_trend.border = THIN_BORDER
        row += 1

    # Trend validation
    trend_dv = DataValidation(type="list", formula1='"↑,↓,→"')
    ws.add_data_validation(trend_dv)
    trend_dv.add('E5:E20')

    # Column widths
    widths = [30, 40, 15, 15, 10]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_vendor_performance_sheet(ws):
    """Create the Vendor Performance sheet."""
    ws.title = "Vendor Performance"

    headers = [
        "Vendor_ID", "Vendor_Name", "Risk_Tier", "Active_Contracts",
        "Total_Deliverables", "Security_Findings_Total", "Critical_Findings",
        "High_Findings", "SLA_Compliance_Rate", "Avg_Remediation_Days",
        "Security_Incidents", "Last_Assessment_Date", "Performance_Score",
        "Performance_Trend", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    tier_dv = DataValidation(type="list", formula1='"Critical,High,Standard"')
    ws.add_data_validation(tier_dv)
    tier_dv.add('C2:C100')

    trend_dv = DataValidation(type="list", formula1='"Improving,Stable,Declining"')
    ws.add_data_validation(trend_dv)
    trend_dv.add('N2:N100')

    # Format input rows
    for row in range(2, 21):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 25, 12, 15, 15, 18, 15, 12, 18, 18, 15, 18, 15, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_exception_register_sheet(ws):
    """Create the Exception Register sheet."""
    ws.title = "Exception Register"

    headers = [
        "Exception_ID", "Exception_Type", "Related_Entity", "Requirement_Reference",
        "Exception_Description", "Risk_Level", "Business_Justification",
        "Compensating_Controls", "Requested_By", "Request_Date", "Approved_By",
        "Approval_Date", "Expiry_Date", "Status", "Renewal_Count",
        "Last_Review_Date", "Next_Review_Date", "Closure_Date", "Closure_Reason"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(type="list", formula1='"Vendor Assessment,Contract Clause,SLA,Testing,Training"')
    ws.add_data_validation(type_dv)
    type_dv.add('B2:B100')

    risk_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(risk_dv)
    risk_dv.add('F2:F100')

    status_dv = DataValidation(type="list", formula1='"Pending,Approved,Rejected,Expired,Renewed"')
    ws.add_data_validation(status_dv)
    status_dv.add('N2:N100')

    closure_dv = DataValidation(type="list", formula1='"Remediated,Risk Accepted,Terminated,N/A"')
    ws.add_data_validation(closure_dv)
    closure_dv.add('S2:S100')

    # Format input rows
    for row in range(2, 31):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 18, 15, 20, 35, 12, 35, 35, 15, 12, 20, 12, 12, 12, 12, 15, 15, 12, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_monitoring_log_sheet(ws):
    """Create the Monitoring Log sheet."""
    ws.title = "Monitoring Log"

    headers = [
        "Log_ID", "Log_Date", "Vendor_ID", "Contract_ID", "Activity_Type",
        "Activity_Description", "Participants", "Findings", "Actions_Required",
        "Action_Owner", "Action_Due_Date", "Action_Status", "Evidence_Reference"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(type="list", formula1='"Status Meeting,Security Review,Audit,Incident Review,Ad-hoc"')
    ws.add_data_validation(type_dv)
    type_dv.add('E2:E100')

    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Complete,Overdue"')
    ws.add_data_validation(status_dv)
    status_dv.add('L2:L100')

    # Format input rows
    for row in range(2, 51):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 12, 12, 12, 18, 40, 25, 35, 35, 20, 15, 15, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_incident_log_sheet(ws):
    """Create the Incident Log sheet."""
    ws.title = "Incident Log"

    headers = [
        "Incident_ID", "Incident_Date", "Vendor_ID", "Contract_ID", "Incident_Type",
        "Severity", "Description", "Root_Cause", "Impact_Assessment",
        "Notification_Date", "Notification_SLA_Met", "Remediation_Actions",
        "Remediation_Date", "Lessons_Learned", "Status", "Closed_Date", "Contract_Impact"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(type="list", formula1='"Data Breach,Vulnerability Exploited,Unauthorized Access,Policy Violation,Other"')
    ws.add_data_validation(type_dv)
    type_dv.add('E2:E100')

    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(severity_dv)
    severity_dv.add('F2:F100')

    sla_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(sla_dv)
    sla_dv.add('K2:K100')

    status_dv = DataValidation(type="list", formula1='"Open,Investigating,Remediated,Closed"')
    ws.add_data_validation(status_dv)
    status_dv.add('O2:O100')

    impact_dv = DataValidation(type="list", formula1='"None,Warning,Review,Suspension,Termination"')
    ws.add_data_validation(impact_dv)
    impact_dv.add('Q2:Q100')

    # Format input rows
    for row in range(2, 31):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 12, 12, 12, 20, 10, 40, 35, 35, 15, 15, 40, 15, 35, 12, 12, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_compliance_score_sheet(ws):
    """Create the Compliance Score Calculation sheet."""
    ws.title = "Compliance Score"

    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="Compliance Score Calculation")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Component", "Weight", "Scoring Criteria", "Data Source", "Raw Score", "Weighted Score"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Score components
    components = [
        ("Vendor Assessment", "20%", "% vendors with current assessment", "Workbook 1"),
        ("Contract Compliance", "25%", "% contracts with all security clauses", "Workbook 2"),
        ("SLA Compliance", "25%", "Weighted avg: Critical (40%) + High (60%)", "Workbook 2"),
        ("Security Testing", "20%", "% deliverables with complete testing", "Workbook 3"),
        ("Exception Management", "10%", "100% - (Overdue exceptions x 10%)", "Sheet 3"),
    ]

    row = 4
    for component, weight, criteria, source in components:
        ws.cell(row=row, column=1, value=component).border = THIN_BORDER
        ws.cell(row=row, column=2, value=weight).border = THIN_BORDER
        ws.cell(row=row, column=3, value=criteria).border = THIN_BORDER
        ws.cell(row=row, column=4, value=source).border = THIN_BORDER
        ws.cell(row=row, column=5).fill = INPUT_FILL
        ws.cell(row=row, column=5).border = THIN_BORDER
        ws.cell(row=row, column=6).fill = LOCKED_FILL
        ws.cell(row=row, column=6).border = THIN_BORDER
        row += 1

    # Total row
    row += 1
    ws.cell(row=row, column=1, value="OVERALL COMPLIANCE SCORE").font = Font(bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER
    ws.cell(row=row, column=6).fill = GREEN_FILL
    ws.cell(row=row, column=6).border = THIN_BORDER
    ws.cell(row=row, column=6).font = Font(bold=True)

    # Score interpretation
    row += 3
    ws.cell(row=row, column=1, value="Score Interpretation:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value=">= 90%: Compliant")
    ws.cell(row=row, column=2).fill = GREEN_FILL
    row += 1
    ws.cell(row=row, column=1, value="70-89%: Needs Improvement")
    ws.cell(row=row, column=2).fill = YELLOW_FILL
    row += 1
    ws.cell(row=row, column=1, value="< 70%: Non-Compliant (Escalation Required)")
    ws.cell(row=row, column=2).fill = RED_FILL

    # Formula note
    row += 2
    ws.cell(row=row, column=1, value="Formula: (Vendor x 0.20) + (Contract x 0.25) + (SLA x 0.25) + (Testing x 0.20) + (Exception x 0.10)")
    ws.cell(row=row, column=1).font = Font(italic=True)

    # Column widths
    widths = [25, 10, 45, 15, 12, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_workbook():
    """Generate the complete assessment workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Create all sheets
    create_instructions_sheet(wb.create_sheet())
    create_executive_dashboard_sheet(wb.create_sheet())
    create_vendor_performance_sheet(wb.create_sheet())
    create_exception_register_sheet(wb.create_sheet())
    create_monitoring_log_sheet(wb.create_sheet())
    create_incident_log_sheet(wb.create_sheet())
    create_compliance_score_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - INITIAL CREATION
# QA_TOOL: Claude Code
# CHANGES: Initial generator creation for A.8.30 Outsourced Development
# =============================================================================
