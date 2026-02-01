#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.30.2 - Contract Compliance
================================================================================

ISO/IEC 27001:2022 Control A.8.30: Outsourced Development
Assessment Domain 2 of 4: Contract Compliance

This script generates a comprehensive Excel assessment workbook for tracking
security clause inclusion in outsourced development contracts and monitoring
ongoing contract compliance.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.30.2"
WORKBOOK_NAME = "Contract Compliance"
CONTROL_ID = "A.8.30"
CONTROL_NAME = "Outsourced Development"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions sheet."""
    ws.title = "Instructions"

    instructions = [
        ["ISMS-IMP-A.8.30.2 - Contract Compliance"],
        [""],
        ["PURPOSE"],
        ["This workbook tracks security clause inclusion in outsourced development contracts"],
        ["and monitors ongoing contract compliance per ISMS-POL-A.8.30."],
        [""],
        ["SHEETS"],
        ["1. Instructions - This guidance sheet"],
        ["2. Contract Inventory - All outsourced development contracts"],
        ["3. Security Clauses - Mandatory security clause verification"],
        ["4. SLA Compliance - Vulnerability remediation SLA tracking"],
        ["5. Subcontractor Approvals - Subcontractor authorization records"],
        ["6. Termination Checklist - Contract closure security verification"],
        [""],
        ["WORKFLOW"],
        ["1. Register new contract in inventory"],
        ["2. Complete security clause checklist"],
        ["3. Obtain security review approval"],
        ["4. Monitor SLA compliance (ongoing)"],
        ["5. Process subcontractor approvals (as needed)"],
        ["6. Complete termination checklist at contract end"],
        [""],
        ["DATA VALIDATION"],
        ["- Yellow cells are input fields"],
        ["- Dropdown lists enforce valid values"],
        ["- Contract_ID must be unique (CTR-XXXX format)"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SHEETS", "WORKFLOW", "DATA VALIDATION"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 80


def create_contract_inventory_sheet(ws):
    """Create the Contract Inventory sheet."""
    ws.title = "Contract Inventory"

    headers = [
        "Contract_ID", "Vendor_ID", "Contract_Name", "Contract_Type",
        "Start_Date", "End_Date", "Project_Classification", "Contract_Value_CHF",
        "Primary_Contact", "Legal_Review_Date", "Security_Review_Date", "Status"
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Add data validation dropdowns
    type_dv = DataValidation(type="list", formula1='"Fixed-price,T&M,Staff Aug,Managed Service"')
    ws.add_data_validation(type_dv)
    type_dv.add('D2:D100')

    class_dv = DataValidation(type="list", formula1='"Critical,High,Standard"')
    ws.add_data_validation(class_dv)
    class_dv.add('G2:G100')

    status_dv = DataValidation(type="list", formula1='"Active,Completed,Terminated"')
    ws.add_data_validation(status_dv)
    status_dv.add('L2:L100')

    # Format input rows
    for row in range(2, 21):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Set column widths
    widths = [12, 12, 35, 18, 12, 12, 18, 15, 25, 15, 18, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_security_clauses_sheet(ws):
    """Create the Security Clauses Checklist sheet."""
    ws.title = "Security Clauses"

    headers = [
        "Contract_ID", "Clause_Category", "Clause_Description", "Included",
        "Clause_Reference", "Modification_Notes", "Reviewed_By", "Review_Date"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Standard security clauses
    clauses = [
        ("Secure Coding Standards", "Compliance with secure coding standards per ISMS-POL-A.8.28"),
        ("Security Policy Compliance", "Adherence to [Organization] security policies"),
        ("Vulnerability Remediation SLAs", "Critical: 7 days, High: 30 days, Medium: 90 days, Low: 180 days"),
        ("Security Testing Rights", "Right to perform security testing on deliverables"),
        ("Audit Rights", "Right to audit vendor security practices"),
        ("Incident Notification", "24-hour notification requirement for security incidents"),
        ("Data Protection", "Confidentiality and data protection requirements"),
        ("Subcontractor Restrictions", "Prior written approval for subcontractors"),
        ("IP/Code Ownership", "Intellectual property and code ownership terms"),
        ("Source Code Escrow", "Escrow arrangement for critical applications"),
        ("Personnel Security", "Background checks and security training requirements"),
        ("Termination Security", "Data return/destruction and access revocation"),
        ("Insurance Requirements", "Cyber liability insurance minimums"),
        ("Liability Provisions", "Liability caps and indemnification"),
    ]

    row = 2
    for category, description in clauses:
        ws.cell(row=row, column=1, value="[Contract_ID]").fill = INPUT_FILL
        ws.cell(row=row, column=2, value=category)
        ws.cell(row=row, column=3, value=description)
        for col in range(4, 9):
            ws.cell(row=row, column=col).fill = INPUT_FILL
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Data validation
    included_dv = DataValidation(type="list", formula1='"Yes,No,N/A,Modified"')
    ws.add_data_validation(included_dv)
    included_dv.add('D2:D100')

    # Column widths
    widths = [12, 25, 55, 12, 18, 35, 20, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_sla_compliance_sheet(ws):
    """Create the SLA Compliance Tracking sheet."""
    ws.title = "SLA Compliance"

    headers = [
        "SLA_ID", "Contract_ID", "Vulnerability_ID", "Severity",
        "Discovery_Date", "SLA_Days", "SLA_Due_Date", "Remediation_Date",
        "SLA_Met", "Exception_Approved", "Exception_Approver", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(severity_dv)
    severity_dv.add('D2:D100')

    sla_days_dv = DataValidation(type="list", formula1='"7,30,90,180"')
    ws.add_data_validation(sla_days_dv)
    sla_days_dv.add('F2:F100')

    met_dv = DataValidation(type="list", formula1='"Met,Missed,Pending,Exception"')
    ws.add_data_validation(met_dv)
    met_dv.add('I2:I100')

    exception_dv = DataValidation(type="list", formula1='"Yes,No,N/A"')
    ws.add_data_validation(exception_dv)
    exception_dv.add('J2:J100')

    # Format input rows
    for row in range(2, 51):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 12, 18, 10, 15, 10, 15, 15, 12, 15, 25, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_subcontractor_approvals_sheet(ws):
    """Create the Subcontractor Approvals sheet."""
    ws.title = "Subcontractor Approvals"

    headers = [
        "Approval_ID", "Contract_ID", "Primary_Vendor_ID", "Subcontractor_Name",
        "Subcontractor_Scope", "Access_Level", "Assessment_Level", "Risk_Classification",
        "Approval_Status", "Approved_By", "Approval_Date", "Expiry_Date",
        "Flow_Down_Verified", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    access_dv = DataValidation(type="list", formula1='"Direct,Via Vendor,None"')
    ws.add_data_validation(access_dv)
    access_dv.add('F2:F100')

    assess_dv = DataValidation(type="list", formula1='"Full,Abbreviated,Vendor Attested"')
    ws.add_data_validation(assess_dv)
    assess_dv.add('G2:G100')

    risk_dv = DataValidation(type="list", formula1='"High,Medium,Low"')
    ws.add_data_validation(risk_dv)
    risk_dv.add('H2:H100')

    status_dv = DataValidation(type="list", formula1='"Approved,Pending,Rejected"')
    ws.add_data_validation(status_dv)
    status_dv.add('I2:I100')

    yn_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('M2:M100')

    # Format input rows
    for row in range(2, 21):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 12, 15, 30, 35, 12, 18, 18, 15, 20, 12, 12, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_termination_checklist_sheet(ws):
    """Create the Termination Checklist sheet."""
    ws.title = "Termination Checklist"

    headers = [
        "Contract_ID", "Termination_Type", "Termination_Date", "Check_Item",
        "Status", "Completion_Date", "Verified_By", "Evidence_Reference"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Standard termination checklist items
    checklist_items = [
        "All access credentials revoked (within 24 hours)",
        "[Organization] data returned or destroyed",
        "Destruction certificate received",
        "Source code transferred (if applicable)",
        "Documentation transferred",
        "Escrow arrangements verified",
        "Outstanding vulnerabilities addressed or risk accepted",
        "Final security review completed",
        "Lessons learned documented",
        "Vendor removed from active registry",
    ]

    row = 2
    for item in checklist_items:
        ws.cell(row=row, column=1, value="[Contract_ID]").fill = INPUT_FILL
        ws.cell(row=row, column=2).fill = INPUT_FILL
        ws.cell(row=row, column=3).fill = INPUT_FILL
        ws.cell(row=row, column=4, value=item)
        for col in range(5, 9):
            ws.cell(row=row, column=col).fill = INPUT_FILL
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Data validation
    type_dv = DataValidation(type="list", formula1='"Completion,Early,Breach"')
    ws.add_data_validation(type_dv)
    type_dv.add('B2:B100')

    status_dv = DataValidation(type="list", formula1='"Complete,Pending,N/A"')
    ws.add_data_validation(status_dv)
    status_dv.add('E2:E100')

    # Column widths
    widths = [12, 15, 15, 50, 12, 15, 20, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_workbook():
    """Generate the complete assessment workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Create all sheets
    create_instructions_sheet(wb.create_sheet())
    create_contract_inventory_sheet(wb.create_sheet())
    create_security_clauses_sheet(wb.create_sheet())
    create_sla_compliance_sheet(wb.create_sheet())
    create_subcontractor_approvals_sheet(wb.create_sheet())
    create_termination_checklist_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - INITIAL CREATION
# QA_TOOL: Claude Code
# CHANGES: Initial generator creation for A.8.30 Outsourced Development
# =============================================================================
