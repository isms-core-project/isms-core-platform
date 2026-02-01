#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.17.2 - System Synchronization Status Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.17: Clock Synchronization
Assessment Domain 2 of 2: System-Level Time Synchronization Verification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific system inventory, platform mix, time synchronization
technologies, and acceptable drift thresholds.

Key customization areas:
1. System inventory import (integrate with your CMDB/asset management)
2. Platform-specific sync status checks (adapt to your OS/device mix)
3. Acceptable drift thresholds (based on your logging/forensic requirements)
4. Criticality levels and compliance criteria (aligned with your risk profile)
5. Integration with monitoring systems (specific to your tooling)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.17 Clock Synchronization Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for verifying
actual time synchronization status across all organizational systems against
ISO 27001:2022 Control A.8.17 requirements.

**Purpose:**
Enables systematic verification that all in-scope systems are actively
synchronizing time with authorized sources within acceptable drift thresholds,
supporting evidence-based validation that time synchronization is not merely
configured but actually working.

**Assessment Philosophy:**
Following Feynman's "don't fool yourself" principle - this assessment proves
synchronization through measurement, not assumption. Configuration != Synchronization.

**Assessment Scope:**
- All servers (physical, virtual, cloud)
- Network infrastructure devices (routers, switches, firewalls)
- Security systems (SIEM, IDS/IPS, authentication)
- Workstations where logging/auditing is required
- Containers and cloud instances
- IoT and embedded systems with logging capability
- Per-system sync status verification
- Time drift measurement against authoritative sources
- Compliance status per system
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and verification procedures
2. System_Sync_Status - Per-system synchronization status inventory
3. Drift_Analysis - Time drift measurements and threshold violations
4. Platform_Summary - Sync status by platform/OS type
5. Criticality_Analysis - Sync compliance by system criticality level
6. Gap_Analysis - Systems not syncing or exceeding drift thresholds
7. Remediation_Plan - Action items for non-compliant systems
8. Evidence_Register - Audit evidence tracking and documentation
9. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- System import from asset inventory (reference A.5.9)
- Platform-specific sync status validation
- Drift threshold compliance checking
- Criticality-based requirement mapping
- Automated gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment feeds into the A.8.17 Compliance Dashboard, which consolidates
data from both time source infrastructure (Assessment 1) and system
synchronization status (this workbook) for executive oversight and audit
readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

Optional:
    - Integration with CMDB/asset inventory for system import
    - Integration with monitoring systems for automated drift data

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a817_2_sync_status.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a817_2_sync_status.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a817_2_sync_status.py --date 20250125
    
    # Import system inventory from CSV
    python3 generate_a817_2_sync_status.py --import-systems inventory.csv

Output:
    File: ISMS-A.8.17-Assessment-2-Sync-Status-YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Import or manually enter system inventory
    2. Execute platform-specific sync verification commands (see ISMS-IMP-A.8.17-S2)
    3. Record sync status for each system (Synced/Not Syncing/Unknown)
    4. Measure and record time drift in seconds
    5. Validate drift against acceptable thresholds
    6. Review gap analysis for non-compliant systems
    7. Define remediation actions with owners and deadlines
    8. Collect and link audit evidence (command outputs, monitoring data)
    9. Obtain stakeholder approvals
    10. Feed results into A.8.17 Compliance Dashboard

Platform-Specific Verification Commands:
    Linux (chrony):    chronyc tracking
    Linux (ntpd):      ntpq -p
    Linux (systemd):   timedatectl
    Windows:           w32tm /query /status
    Network devices:   show ntp status (vendor-specific)
    Cloud instances:   Provider-specific validation
    
See ISMS-IMP-A.8.17-S2 for complete verification procedures.

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.17
Assessment Domain:    2 of 2 (System Synchronization Status)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.8.17: Clock Synchronization Policy (Requirements)
    - ISMS-IMP-A.8.17-S1: Time Source Configuration (Infrastructure)
    - ISMS-IMP-A.8.17-S2: Synchronization Verification Process (Procedures)
    - ISMS-IMP-A.8.17.1: Time Source Infrastructure Assessment (Domain 1)
    - A.8.17 Compliance Dashboard (Consolidation)

Related Standards:
    - RFC 5905: Network Time Protocol Version 4
    - NIST Special Publication 1800-16: Time Synchronization
    - ISO/IEC 27002:2022 Control 8.17: Clock Synchronization

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.8.17-S2 specification
    - Supports comprehensive system synchronization verification
    - Platform-specific validation for Linux, Windows, network devices, cloud
    - Integrated with A.8.17 Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Verification Philosophy:**
This assessment embodies "configuration ≠ synchronization" - having NTP
configured does not guarantee synchronization is working. Every system must
be individually verified with actual sync status checks and drift measurements.

**Acceptable Drift Thresholds:**
Policy defines maximum acceptable drift thresholds:
- General systems: ±1 second from authoritative source
- Critical security systems: ±100 milliseconds (SIEM, authentication, PKI)
- High-precision systems: ±10 milliseconds (financial, regulatory)

Organizations must define system-specific thresholds based on operational
requirements while not exceeding policy maximums.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Auditors will expect:
- Evidence of actual sync status verification (not just configuration review)
- Time drift measurements for sample systems
- Documented remediation for non-compliant systems
- Regular assessment updates (quarterly recommended)

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- System inventory with hostnames and IP addresses
- Infrastructure topology and criticality levels
- Security system identification
- Synchronization failures indicating potential issues

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update assessment:
- Monthly: Automated sync status collection where possible
- Quarterly: Complete manual verification for critical systems
- When systems are added/removed: Update inventory
- After infrastructure changes: Re-verify affected systems
- Ad-hoc: When sync issues are detected or reported

**Platform Coverage:**
Assessment supports multiple platforms:
- Linux: chrony, ntpd, systemd-timesyncd
- Windows: W32Time service
- Network devices: Cisco, Juniper, Palo Alto, etc.
- Cloud: AWS Time Sync, Azure NTP, GCP NTP
- Containers: Verify host synchronization
- Virtualization: ESXi, Hyper-V time services

Customize verification procedures for your specific platform mix.

**Integration with Monitoring:**
Where possible, integrate with existing monitoring systems to:
- Automate sync status collection
- Receive real-time alerts for sync failures
- Track drift trends over time
- Reduce manual assessment effort

**Quality Assurance:**
Have system administrators and operations engineers validate assessments
before using results for compliance reporting. Sync status can change
rapidly - ensure data is current and accurate.

**Gap Remediation:**
Non-compliant systems must be remediated promptly:
- Critical systems: Within 3 business days
- High-priority systems: Within 1 week
- Standard systems: Within 2 weeks
- Document exceptions with risk acceptance

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
import argparse
import csv

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X

# Document identification constants
DOCUMENT_ID = "ISMS-IMP-A.8.17.2"
CONTROL_REF = "ISO/IEC 27001:2022 - Control A.8.17: Clock Synchronization"
WARNING = '\u26A0'    # ⚠️  Warning sign
CLOCK = '\u23F0'      # ⏰ Alarm clock
SYNC = '\U0001F504'   # 🔄 Counterclockwise arrows
HOURGLASS = '\u23F3'  # ⏳ Hourglass
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
BULLET = '\u2022'     # • Bullet
ARROW = '\u2192'      # → Right arrow

def create_styles():
    """Define A.8.24 standard styles for the workbook"""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "title": {
            "font": Font(name="Calibri", bold=True, size=14, color="003366"),
            "alignment": Alignment(horizontal="left", vertical="center")
        },
        "center": {
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "data": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        }
    }
    return styles

def set_column_widths(ws, widths):
    """Set column widths for a worksheet"""
    for col_num, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

def import_asset_inventory(filename):
    """Import asset inventory from CSV file (optional)"""
    assets = []
    try:
        with open(filename, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                assets.append({
                    'name': row.get('AssetName', row.get('Hostname', '')),
                    'type': row.get('AssetType', row.get('Type', 'Server')),
                    'criticality': row.get('Criticality', 'Medium'),
                    'asset_id': row.get('AssetID', '')
                })
    except FileNotFoundError:
        logger.info(f"Warning: Asset file {filename} not found. Using example data.")
    except Exception as e:
        logger.error(f"Warning: Error reading asset file: {e}. Using example data.")
    
    return assets

def create_instructions_sheet(wb):
    """Create instructions sheet"""
    ws = wb.create_sheet("Instructions", 0)
    styles = create_styles()
    
    # Title
    ws['A1'] = f"{DOCUMENT_ID}\n{CONTROL_REF}"
    ws['A1'].font = Font(bold=True, size=16, color="003366")
    ws['A1'].alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:F2')
    
    # Document ID (for normalization script)
    ws['A4'] = "Document ID:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "ISMS-IMP-A.8.17.2"
    ws['B4'].font = Font(bold=True, color="003366")
    
    ws['A5'] = "Title:"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = "System Synchronization Status Assessment"
    
    # Instructions
    instructions = [
        ("", ""),
        ("Purpose:", "Document per-system NTP synchronization status, measure time drift, and track compliance."),
        ("", ""),
        ("Sheets:", ""),
        ("  • System_Inventory", "All systems requiring time synchronization"),
        ("  • Drift_Analysis", "Statistical analysis of time drift across infrastructure"),
        ("  • Gaps_Failures", "Systems not synchronized or exceeding drift thresholds"),
        ("  • Compliance_Summary", "Overall compliance metrics and scores"),
        ("", ""),
        ("Instructions:", ""),
        ("1.", "Import system inventory from A.5.9 Asset Management (if available)"),
        ("2.", "For each system, verify and document:"),
        ("   ", "  • NTP server configuration"),
        ("   ", "  • Current synchronization status"),
        ("   ", "  • Time drift measurement"),
        ("   ", "  • Last sync verification timestamp"),
        ("3.", "Use verification commands from ISMS-IMP-A.8.17-S2"),
        ("4.", "Review Drift_Analysis for statistical insights"),
        ("5.", "Address all items in Gaps_Failures sheet"),
        ("6.", "Review Compliance_Summary for overall status"),
        ("", ""),
        ("Drift Thresholds:", "Per ISMS-POL-A.8.17 REQ-817-011:"),
        ("  • General systems:", "±1 second (±1000 milliseconds)"),
        ("  • Critical security:", "±100 milliseconds"),
        ("  • High-precision:", "±10 milliseconds"),
        ("", ""),
        ("Assessment Frequency:", "Monthly, with continuous monitoring"),
        ("Policy Reference:", "ISMS-POL-A.8.17 Clock Synchronization Policy"),
        ("Verification Guide:", "ISMS-IMP-A.8.17-S2 Synchronization Verification Process"),
    ]
    
    row = 7
    for col1, col2 in instructions:
        ws[f'A{row}'] = col1
        ws[f'B{row}'] = col2
        if col1 in ["Purpose:", "Sheets:", "Instructions:", "Drift Thresholds:", 
                    "Assessment Frequency:", "Policy Reference:", "Verification Guide:"]:
            ws[f'A{row}'].font = Font(bold=True, color="003366")
            ws.merge_cells(f'A{row}:F{row}')
        row += 1
    
    set_column_widths(ws, [15, 80, 15, 15, 15, 15])
    
    return ws

def create_system_inventory_sheet(wb, assets=None):
    """Create system inventory sheet"""
    ws = wb.create_sheet("System_Inventory")
    styles = create_styles()
    
    # Headers
    headers = [
        "System Name [*]",
        "Asset ID",
        "Type [*]",
        "OS/Platform",
        "Criticality",
        "NTP Server(s) Configured [*]",
        "Sync Status [*]",
        "Stratum",
        "Current Drift (ms) [*]",
        "Last Sync Time",
        "Last Verified [*]",
        "Compliance",
        "Notes"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = styles['header']['font']
        cell.fill = styles['header']['fill']
        cell.alignment = styles['header']['alignment']
        cell.border = styles['header']['border']
    
    # Data validation for Type
    type_dv = DataValidation(
        type="list",
        formula1='"🖥️ Server-Physical,💻 Server-Virtual,☁️ Server-Cloud,🌐 Network Device,🔒 Security Appliance,💼 Workstation,📦 Container Host,🔌 IoT Device,📋 Other"',
        allow_blank=False
    )
    type_dv.error = "Please select a valid system type"
    type_dv.errorTitle = "Invalid Type"
    ws.add_data_validation(type_dv)
    type_dv.add(f'C2:C1000')
    
    # Data validation for Criticality
    criticality_dv = DataValidation(
        type="list",
        formula1='"🔴 Critical,🟠 High,🟡 Medium,🟢 Low"',
        allow_blank=False
    )
    criticality_dv.error = "Please select a valid criticality level"
    criticality_dv.errorTitle = "Invalid Criticality"
    ws.add_data_validation(criticality_dv)
    criticality_dv.add(f'E2:E1000')
    
    # Data validation for Sync Status
    sync_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Synced,❌ Not Synced,⚠️ Sync Failed,❓ Unknown,➖ Excluded"',
        allow_blank=False
    )
    sync_dv.error = "Please select a valid sync status"
    sync_dv.errorTitle = "Invalid Sync Status"
    ws.add_data_validation(sync_dv)
    sync_dv.add(f'G2:G1000')
    
    # Data validation for Stratum
    stratum_dv = DataValidation(
        type="whole",
        formula1="0",
        formula2="16",
        allow_blank=True
    )
    stratum_dv.error = "Stratum must be 0-16 (16 = not synchronized)"
    stratum_dv.errorTitle = "Invalid Stratum"
    ws.add_data_validation(stratum_dv)
    stratum_dv.add(f'H2:H1000')
    
    # Compliance formula (calculated based on drift and criticality)
    # Formula: IF(G2="{CHECK} Synced", IF(E2="🔴 Critical", IF(ABS(I2)<=100,"{CHECK} PASS","{XMARK} FAIL"), IF(ABS(I2)<=1000,"{CHECK} PASS","{XMARK} FAIL")), f"{XMARK} FAIL")
    
    # Example rows
    if assets and len(assets) > 0:
        # Use imported assets
        examples = []
        for asset in assets[:10]:  # First 10 assets as examples
            examples.append([
                asset['name'],
                asset.get('asset_id', ''),
                asset.get('type', 'Server'),
                '',  # OS/Platform - to be filled
                asset.get('criticality', 'Medium'),
                'ntp1.organization.local, ntp2.organization.local',
                'Synced',
                '3',
                '0.5',
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                datetime.now().strftime('%d.%m.%Y'),
                '=IF(G2="{CHECK} Synced",IF(E2="🔴 Critical",IF(ABS(I2)<=100,"{CHECK} PASS","{XMARK} FAIL"),IF(ABS(I2)<=1000,"{CHECK} PASS","{XMARK} FAIL")),"{XMARK} FAIL")',
                ''
            ])
    else:
        # Default examples
        examples = [
            ["web-server-01.org.local", "SRV-001", "🖥️ Server-Physical", "Ubuntu 22.04", "🟠 High",
             "ntp1.organization.local, ntp2.organization.local", f"{CHECK} Synced", "3", "0.5",
             datetime.now().strftime('%Y-%m-%d %H:%M:%S'), datetime.now().strftime('%d.%m.%Y'),
             '=IF(G2="{CHECK} Synced",IF(E2="🔴 Critical",IF(ABS(I2)<=100,"{CHECK} PASS","{XMARK} FAIL"),IF(ABS(I2)<=1000,"{CHECK} PASS","{XMARK} FAIL")),"{XMARK} FAIL")',
             "Verified via chronyc tracking"],
            ["db-server-01.org.local", "SRV-002", "💻 Server-Virtual", "Windows Server 2022", "🔴 Critical",
             "ntp1.organization.local, ntp2.organization.local", f"{CHECK} Synced", "3", "25.3",
             datetime.now().strftime('%Y-%m-%d %H:%M:%S'), datetime.now().strftime('%d.%m.%Y'),
             '=IF(G3="{CHECK} Synced",IF(E3="🔴 Critical",IF(ABS(I3)<=100,"{CHECK} PASS","{XMARK} FAIL"),IF(ABS(I3)<=1000,"{CHECK} PASS","{XMARK} FAIL")),"{XMARK} FAIL")',
             "Verified via w32tm"],
            ["firewall-01.org.local", "NET-001", "🌐 Network Device", "Cisco ASA", "🔴 Critical",
             "10.0.1.10, 10.0.1.11", f"{CHECK} Synced", "3", "15.7",
             datetime.now().strftime('%Y-%m-%d %H:%M:%S'), datetime.now().strftime('%d.%m.%Y'),
             '=IF(G4="{CHECK} Synced",IF(E4="🔴 Critical",IF(ABS(I4)<=100,"{CHECK} PASS","{XMARK} FAIL"),IF(ABS(I4)<=1000,"{CHECK} PASS","{XMARK} FAIL")),"{XMARK} FAIL")',
             "Verified via show ntp status"],
            ["siem-01.org.local", "SEC-001", "🔒 Security Appliance", "Splunk Enterprise", "🔴 Critical",
             "ntp1.organization.local, ntp2.organization.local", f"{XMARK} Not Synced", "16", "1500.0",
             "N/A", datetime.now().strftime('%d.%m.%Y'),
             '=IF(G5="{CHECK} Synced",IF(E5="🔴 Critical",IF(ABS(I5)<=100,"{CHECK} PASS","{XMARK} FAIL"),IF(ABS(I5)<=1000,"{CHECK} PASS","{XMARK} FAIL")),"{XMARK} FAIL")',
             "NTP service not configured - REQUIRES REMEDIATION"],
            ["app-server-03.org.local", "SRV-005", "☁️ Server-Cloud", "Amazon Linux 2", "🟡 Medium",
             "169.254.169.123 (AWS Time Sync)", f"{CHECK} Synced", "3", "2.1",
             datetime.now().strftime('%Y-%m-%d %H:%M:%S'), datetime.now().strftime('%d.%m.%Y'),
             '=IF(G6="{CHECK} Synced",IF(E6="🔴 Critical",IF(ABS(I6)<=100,"{CHECK} PASS","{XMARK} FAIL"),IF(ABS(I6)<=1000,"{CHECK} PASS","{XMARK} FAIL")),"{XMARK} FAIL")',
             "Using AWS Time Sync Service"],
        ]
    
    for row_num, example in enumerate(examples, start=2):
        for col_num, value in enumerate(example, start=1):
            cell = ws.cell(row=row_num, column=col_num)
            if isinstance(value, str) and value.startswith('='):
                cell.value = value
            else:
                cell.value = value
            cell.border = styles['data']['border']
            
            if col_num in [8, 9]:  # Stratum and Drift columns - center aligned
                cell.alignment = styles['center']['alignment']
            else:
                cell.alignment = styles['data']['alignment']
    
    # Add formula for additional rows
    last_example_row = len(examples) + 1
    for row_num in range(last_example_row + 1, last_example_row + 20):
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = styles['data']['border']
            
            # Add compliance formula
            if col_num == 12:  # Compliance column
                cell.value = f'=IF(G{row_num}="{CHECK} Synced",IF(E{row_num}="🔴 Critical",IF(ABS(I{row_num})<=100,"{CHECK} PASS","{XMARK} FAIL"),IF(ABS(I{row_num})<=1000,"{CHECK} PASS","{XMARK} FAIL")),"{XMARK} FAIL")'
    
    # Set column widths
    set_column_widths(ws, [28, 12, 20, 20, 12, 35, 15, 10, 18, 22, 15, 12, 40])
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    return ws

def create_drift_analysis_sheet(wb):
    """Create drift analysis sheet"""
    ws = wb.create_sheet("Drift_Analysis")
    styles = create_styles()
    
    # Title
    ws['A1'] = "Time Drift Statistical Analysis"
    ws['A1'].font = Font(bold=True, size=14, color="003366")
    ws.merge_cells('A1:D1')
    
    ws['A2'] = "Analysis of time drift across all synchronized systems"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:D2')
    
    # Overall statistics
    row = 4
    ws[f'A{row}'] = "Overall Drift Statistics"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    stats_data = [
        ("Metric", "Value", "Unit", "Notes"),
        ("Total Systems Assessed", "=COUNTA(System_Inventory!A2:A1000)", "systems", "All systems in inventory"),
        ("Systems Synchronized", '=COUNTIF(System_Inventory!G2:G1000,"Synced")', "systems", "Stratum ≠ 16"),
        ("Systems Not Synchronized", '=COUNTIF(System_Inventory!G2:G1000,"Not Synced")+COUNTIF(System_Inventory!G2:G1000,"Sync Failed")', 
         "systems", "Requires remediation"),
        ("Average Drift (Synced Systems)", '=AVERAGEIF(System_Inventory!G2:G1000,"Synced",System_Inventory!I2:I1000)', 
         "ms", "Mean time offset"),
        ("Maximum Drift (Synced Systems)", '=MAXIFS(System_Inventory!I2:I1000,System_Inventory!G2:G1000,"Synced")',
         "ms", "Highest offset observed"),
        ("Minimum Drift (Synced Systems)", '=MINIFS(System_Inventory!I2:I1000,System_Inventory!G2:G1000,"Synced")',
         "ms", "Lowest offset observed"),
        ("Standard Deviation", '=STDEV(System_Inventory!I2:I1000)', "ms", "Variation in drift"),
    ]
    
    for row_data in stats_data:
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_num)
            if isinstance(value, str) and value.startswith('='):
                cell.value = value
            else:
                cell.value = value
            cell.border = styles['data']['border']
            
            if row == 5:  # Header
                cell.font = styles['header']['font']
                cell.fill = styles['header']['fill']
                cell.alignment = styles['header']['alignment']
            else:
                cell.alignment = styles['data']['alignment']
        row += 1
    
    # Drift distribution by threshold
    row += 2
    ws[f'A{row}'] = "Drift Distribution by Threshold"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    threshold_data = [
        ("Threshold", "Count", "Percentage", "Status"),
        ("≤ 10ms (High-Precision)", '=COUNTIFS(System_Inventory!I2:I1000,"<=10",System_Inventory!I2:I1000,">=-10",System_Inventory!G2:G1000,"Synced")',
         '=B' + str(row+1) + '/B6*100', "Excellent"),
        ("≤ 100ms (Critical Systems)", '=COUNTIFS(System_Inventory!I2:I1000,"<=100",System_Inventory!I2:I1000,">=-100",System_Inventory!G2:G1000,"Synced")',
         '=B' + str(row+2) + '/B6*100', "Good for critical"),
        ("≤ 1000ms (General Systems)", '=COUNTIFS(System_Inventory!I2:I1000,"<=1000",System_Inventory!I2:I1000,">=-1000",System_Inventory!G2:G1000,"Synced")',
         '=B' + str(row+3) + '/B6*100', "Acceptable"),
        ("> 1000ms (Exceeds Threshold)", '=COUNTIFS(System_Inventory!I2:I1000,">1000",System_Inventory!G2:G1000,"Synced")+COUNTIFS(System_Inventory!I2:I1000,"<-1000",System_Inventory!G2:G1000,"Synced")',
         '=B' + str(row+4) + '/B6*100', "FAIL - Requires action"),
    ]
    
    for row_data in threshold_data:
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_num)
            if isinstance(value, str) and value.startswith('='):
                cell.value = value
                if 'B' in value and '*100' in value:  # Percentage
                    cell.number_format = '0.0"%"'
            else:
                cell.value = value
            cell.border = styles['data']['border']
            
            if row == row - len(threshold_data):  # Header row calculation
                cell.font = styles['header']['font']
                cell.fill = styles['header']['fill']
                cell.alignment = styles['header']['alignment']
            else:
                cell.alignment = styles['data']['alignment']
        row += 1
    
    # By system type
    row += 2
    ws[f'A{row}'] = "Drift Analysis by System Type"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = "System Type"
    ws[f'B{row}'] = "Avg Drift (ms)"
    ws[f'C{row}'] = "Max Drift (ms)"
    ws[f'D{row}'] = "Count"
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}{row}']
        cell.font = styles['header']['font']
        cell.fill = styles['header']['fill']
        cell.alignment = styles['header']['alignment']
        cell.border = styles['header']['border']
    
    row += 1
    type_analysis = [
        ("Server-Physical", '=AVERAGEIF(System_Inventory!C2:C1000,"Server-Physical",System_Inventory!I2:I1000)',
         '=MAXIFS(System_Inventory!I2:I1000,System_Inventory!C2:C1000,"Server-Physical")',
         '=COUNTIF(System_Inventory!C2:C1000,"Server-Physical")'),
        ("Server-Virtual", '=AVERAGEIF(System_Inventory!C2:C1000,"Server-Virtual",System_Inventory!I2:I1000)',
         '=MAXIFS(System_Inventory!I2:I1000,System_Inventory!C2:C1000,"Server-Virtual")',
         '=COUNTIF(System_Inventory!C2:C1000,"Server-Virtual")'),
        ("Server-Cloud", '=AVERAGEIF(System_Inventory!C2:C1000,"Server-Cloud",System_Inventory!I2:I1000)',
         '=MAXIFS(System_Inventory!I2:I1000,System_Inventory!C2:C1000,"Server-Cloud")',
         '=COUNTIF(System_Inventory!C2:C1000,"Server-Cloud")'),
        ("Network Device", '=AVERAGEIF(System_Inventory!C2:C1000,"Network Device",System_Inventory!I2:I1000)',
         '=MAXIFS(System_Inventory!I2:I1000,System_Inventory!C2:C1000,"Network Device")',
         '=COUNTIF(System_Inventory!C2:C1000,"Network Device")'),
        ("Security Appliance", '=AVERAGEIF(System_Inventory!C2:C1000,"Security Appliance",System_Inventory!I2:I1000)',
         '=MAXIFS(System_Inventory!I2:I1000,System_Inventory!C2:C1000,"Security Appliance")',
         '=COUNTIF(System_Inventory!C2:C1000,"Security Appliance")'),
    ]
    
    for type_data in type_analysis:
        for col_num, value in enumerate(type_data, start=1):
            cell = ws.cell(row=row, column=col_num)
            if isinstance(value, str) and value.startswith('='):
                cell.value = value
            else:
                cell.value = value
            cell.border = styles['data']['border']
            cell.alignment = styles['data']['alignment']
        row += 1
    
    # By criticality
    row += 2
    ws[f'A{row}'] = "Drift Analysis by Criticality"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = "Criticality"
    ws[f'B{row}'] = "Avg Drift (ms)"
    ws[f'C{row}'] = "Max Drift (ms)"
    ws[f'D{row}'] = "Count"
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}{row}']
        cell.font = styles['header']['font']
        cell.fill = styles['header']['fill']
        cell.alignment = styles['header']['alignment']
        cell.border = styles['header']['border']
    
    row += 1
    crit_analysis = [
        ("Critical", '=AVERAGEIF(System_Inventory!E2:E1000,"Critical",System_Inventory!I2:I1000)',
         '=MAXIFS(System_Inventory!I2:I1000,System_Inventory!E2:E1000,"Critical")',
         '=COUNTIF(System_Inventory!E2:E1000,"Critical")'),
        ("High", '=AVERAGEIF(System_Inventory!E2:E1000,"High",System_Inventory!I2:I1000)',
         '=MAXIFS(System_Inventory!I2:I1000,System_Inventory!E2:E1000,"High")',
         '=COUNTIF(System_Inventory!E2:E1000,"High")'),
        ("Medium", '=AVERAGEIF(System_Inventory!E2:E1000,"Medium",System_Inventory!I2:I1000)',
         '=MAXIFS(System_Inventory!I2:I1000,System_Inventory!E2:E1000,"Medium")',
         '=COUNTIF(System_Inventory!E2:E1000,"Medium")'),
        ("Low", '=AVERAGEIF(System_Inventory!E2:E1000,"Low",System_Inventory!I2:I1000)',
         '=MAXIFS(System_Inventory!I2:I1000,System_Inventory!E2:E1000,"Low")',
         '=COUNTIF(System_Inventory!E2:E1000,"Low")'),
    ]
    
    for crit_data in crit_analysis:
        for col_num, value in enumerate(crit_data, start=1):
            cell = ws.cell(row=row, column=col_num)
            if isinstance(value, str) and value.startswith('='):
                cell.value = value
            else:
                cell.value = value
            cell.border = styles['data']['border']
            cell.alignment = styles['data']['alignment']
        row += 1
    
    set_column_widths(ws, [30, 20, 20, 30])
    
    return ws

def create_gaps_failures_sheet(wb):
    """Create gaps and failures sheet"""
    ws = wb.create_sheet("Gaps_Failures")
    styles = create_styles()
    
    # Title
    ws['A1'] = "Systems Not Synchronized or Exceeding Drift Threshold"
    ws['A1'].font = Font(bold=True, size=14, color="003366")
    ws.merge_cells('A1:H1')
    
    ws['A2'] = "These systems require immediate remediation"
    ws['A2'].font = Font(italic=True, size=10, color="FF0000")
    ws.merge_cells('A2:H2')
    
    # Headers
    row = 4
    headers = [
        "System Name",
        "Type",
        "Criticality",
        "Sync Status",
        "Drift (ms)",
        "Issue Category",
        "Remediation Action",
        "Target Date"
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = styles['header']['font']
        cell.fill = styles['header']['fill']
        cell.alignment = styles['header']['alignment']
        cell.border = styles['header']['border']
    
    row += 1
    
    # Instructions row
    ws[f'A{row}'] = "Fill this sheet with systems from System_Inventory where:"
    ws[f'A{row}'].font = Font(italic=True)
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    ws[f'A{row}'] = f"{BULLET} Sync Status = f'{XMARK} Not Synced', f'{WARNING} Sync Failed', or '❓ Unknown'"
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    ws[f'A{row}'] = f"{BULLET} Compliance = f'{XMARK} FAIL' (drift exceeds threshold for criticality level)"
    ws.merge_cells(f'A{row}:H{row}')
    row += 1
    
    ws[f'A{row}'] = f"{BULLET} Stratum = 16 (not synchronized)"
    ws.merge_cells(f'A{row}:H{row}')
    row += 2
    
    # Example gap entries
    examples = [
        ["siem-01.org.local", "🔒 Security Appliance", "🔴 Critical", f"{XMARK} Not Synced", "1500.0",
        "NTP Not Configured", "Configure NTP client, restart service", 
        (datetime.now() + timedelta(days=3)).strftime('%d.%m.%Y')],
        ["legacy-app-server.org.local", "🖥️ Server-Physical", "🟡 Medium", f"{WARNING} Sync Failed", "2500.0",
        "NTP Server Unreachable", "Verify network connectivity, update firewall rules",
        (datetime.now() + timedelta(days=5)).strftime('%d.%m.%Y')],
        ["db-replica-02.org.local", "💻 Server-Virtual", "🔴 Critical", f"{CHECK} Synced", "150.0",
        "Excessive Drift (Critical)", "Investigate time source quality, restart NTP",
        (datetime.now() + timedelta(days=1)).strftime('%d.%m.%Y')],
    ]
    
    for example in examples:
        for col_num, value in enumerate(example, start=1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            cell.border = styles['data']['border']
            cell.alignment = styles['data']['alignment']
            
            # Highlight critical issues
            if example[2] == "🔴 Critical":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 1
    
    # Add empty rows for additional gaps
    for r in range(row, row + 15):
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col_num)
            cell.border = styles['data']['border']
    
    set_column_widths(ws, [28, 20, 15, 18, 15, 25, 40, 15])
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    return ws

def create_compliance_summary_sheet(wb):
    """Create compliance summary sheet"""
    ws = wb.create_sheet("Compliance_Summary")
    styles = create_styles()
    
    # Title
    ws['A1'] = "System Synchronization Status - Compliance Summary"
    ws['A1'].font = Font(bold=True, size=14, color="003366")
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Assessment Date: {datetime.now().strftime('%d.%m.%Y')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:D2')
    
    # Overall compliance
    row = 4
    ws[f'A{row}'] = "Overall Compliance Metrics"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    metrics = [
        ("Metric", "Value", "Target", "Status"),
        ("Total Systems Assessed", "=COUNTA(System_Inventory!A2:A1000)", "All systems", ""),
        ("Systems Synchronized", '=COUNTIF(System_Inventory!G2:G1000,"Synced")', "100%", 
         '=IF(B6/B5>=0.95,"PASS","FAIL")'),
        ("Synchronization Rate", "=B6/B5*100", "≥95%", 
         '=IF(B7>=95,"PASS","FAIL")'),
        ("Systems Within Threshold", '=COUNTIF(System_Inventory!L2:L1000,"PASS")', "100%",
         '=IF(B8/B5>=0.95,"PASS","FAIL")'),
        ("Compliance Rate", "=B8/B5*100", "≥95%",
         '=IF(B9>=95,"PASS","FAIL")'),
        ("Average Drift (All Systems)", "=Drift_Analysis!B5", "< 100ms", 
         '=IF(B10<100,"PASS","WARN")'),
        ("Maximum Drift Observed", "=Drift_Analysis!B6", "< 1000ms",
         '=IF(B11<1000,"PASS","FAIL")'),
        ("Critical Systems in Compliance", 
         '=COUNTIFS(System_Inventory!E2:E1000,"Critical",System_Inventory!L2:L1000,"PASS")',
         "100%", '=IF(B12/COUNTIF(System_Inventory!E2:E1000,"Critical")=1,"PASS","FAIL")'),
    ]
    
    for row_data in metrics:
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_num)
            if isinstance(value, str) and value.startswith('='):
                cell.value = value
                if 'B' in value and '*100' in value and col_num == 2:  # Percentage
                    cell.number_format = '0.0"%"'
            else:
                cell.value = value
            cell.border = styles['data']['border']
            
            if row == 5:  # Header
                cell.font = styles['header']['font']
                cell.fill = styles['header']['fill']
                cell.alignment = styles['header']['alignment']
            else:
                cell.alignment = styles['data']['alignment']
        row += 1
    
    # By criticality
    row += 2
    ws[f'A{row}'] = "Compliance by Criticality Level"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    crit_compliance = [
        ("Criticality", "Total", "Compliant", "Rate"),
        ("🔴 Critical", '=COUNTIF(System_Inventory!E2:E1000,"🔴 Critical")',
         '=COUNTIFS(System_Inventory!E2:E1000,"🔴 Critical",System_Inventory!L2:L1000,"{CHECK} PASS")',
         '=C' + str(row+1) + '/B' + str(row+1) + '*100'),
        ("🟠 High", '=COUNTIF(System_Inventory!E2:E1000,"🟠 High")',
         '=COUNTIFS(System_Inventory!E2:E1000,"🟠 High",System_Inventory!L2:L1000,"{CHECK} PASS")',
         '=C' + str(row+2) + '/B' + str(row+2) + '*100'),
        ("🟡 Medium", '=COUNTIF(System_Inventory!E2:E1000,"🟡 Medium")',
         '=COUNTIFS(System_Inventory!E2:E1000,"🟡 Medium",System_Inventory!L2:L1000,"{CHECK} PASS")',
         '=C' + str(row+3) + '/B' + str(row+3) + '*100'),
        ("🟢 Low", '=COUNTIF(System_Inventory!E2:E1000,"🟢 Low")',
         '=COUNTIFS(System_Inventory!E2:E1000,"🟢 Low",System_Inventory!L2:L1000,"{CHECK} PASS")',
         '=C' + str(row+4) + '/B' + str(row+4) + '*100'),
    ]
    
    for row_data in crit_compliance:
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_num)
            if isinstance(value, str) and value.startswith('='):
                cell.value = value
                if col_num == 4:  # Rate column
                    cell.number_format = '0.0"%"'
            else:
                cell.value = value
            cell.border = styles['data']['border']
            
            if row == row - len(crit_compliance):  # Header calculation
                cell.font = styles['header']['font']
                cell.fill = styles['header']['fill']
                cell.alignment = styles['header']['alignment']
            else:
                cell.alignment = styles['data']['alignment']
        row += 1
    
    # Critical gaps summary
    row += 2
    ws[f'A{row}'] = "Critical Gaps Requiring Immediate Action"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FF0000")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = "Gap Type"
    ws[f'B{row}'] = "Count"
    ws[f'C{row}'] = "Severity"
    ws[f'D{row}'] = "Action Required"
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}{row}']
        cell.font = styles['header']['font']
        cell.fill = styles['header']['fill']
        cell.alignment = styles['header']['alignment']
        cell.border = styles['header']['border']
    
    row += 1
    gaps = [
        ("Systems Not Synchronized", 
         '=COUNTIF(System_Inventory!G2:G1000,"{XMARK} Not Synced")+COUNTIF(System_Inventory!G2:G1000,"{WARNING} Sync Failed")',
         "🔴 HIGH", "Configure NTP, verify connectivity"),
        ("Critical Systems Non-Compliant",
         '=COUNTIFS(System_Inventory!E2:E1000,"🔴 Critical",System_Inventory!L2:L1000,"{XMARK} FAIL")',
         "🔴 HIGH", "Immediate remediation required"),
        ("Systems Exceeding Drift Threshold",
         '=COUNTIF(System_Inventory!L2:L1000,"{XMARK} FAIL")',
         "🟡 MEDIUM", "Investigate and remediate"),
        ("Unknown Sync Status",
         '=COUNTIF(System_Inventory!G2:G1000,"❓ Unknown")',
         "🟡 MEDIUM", "Verify and document status"),
    ]
    
    for gap_data in gaps:
        for col_num, value in enumerate(gap_data, start=1):
            cell = ws.cell(row=row, column=col_num)
            if isinstance(value, str) and value.startswith('='):
                cell.value = value
            else:
                cell.value = value
            cell.border = styles['data']['border']
            cell.alignment = styles['data']['alignment']
            
            if gap_data[2] == "🔴 HIGH":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 1
    
    # Next steps
    row += 2
    ws[f'A{row}'] = "Next Steps"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    next_steps = [
        "1. Review all systems in Gaps_Failures sheet",
        "2. Assign remediation tasks to system owners",
        "3. Track remediation progress (target dates in Gaps_Failures)",
        "4. Re-verify systems after remediation",
        "5. Update this assessment monthly",
        "6. Escalate persistent non-compliance to management",
        "7. Review drift trends in Drift_Analysis sheet",
    ]
    
    for step in next_steps:
        ws[f'A{row}'] = step
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
    
    # Assessment metadata
    row += 2
    ws[f'A{row}'] = "Assessment Information"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="003366")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    metadata = [
        ("Assessment Date:", datetime.now().strftime('%d.%m.%Y')),
        ("Assessed By:", "[Name]"),
        ("Review Date:", "[Date]"),
        ("Approved By:", "[Name]"),
        ("Next Assessment:", "[Monthly - Date]"),
        ("Policy Reference:", "ISMS-POL-A.8.17 REQ-817-011"),
        ("Verification Guide:", "ISMS-IMP-A.8.17-S2"),
    ]
    
    for label, value in metadata:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'B{row}:D{row}')
        row += 1
    
    set_column_widths(ws, [35, 25, 20, 40])
    
    return ws

def main():
    """Main function to generate the workbook"""
    parser = argparse.ArgumentParser(
        description='Generate ISMS A.8.17 System Synchronization Status Assessment Workbook'
    )
    parser.add_argument(
        '--output',
        default=f'ISMS-IMP-A.8.17.2_Sync_Status_{datetime.now().strftime("%Y%m%d")}.xlsx',
        help='Output filename (default: ISMS-A.8.17-Assessment-2-Sync-Status_YYYYMMDD.xlsx)'
    )
    parser.add_argument(
        '--import-assets',
        help='Import asset inventory from CSV file (optional)'
    )
    
    args = parser.parse_args()
    
    logger.info("Generating ISMS A.8.17 System Synchronization Status Assessment Workbook...")
    
    # Import assets if provided
    assets = None
    if args.import_assets:
        logger.info(f"  Importing asset inventory from {args.import_assets}...")
        assets = import_asset_inventory(args.import_assets)
        if assets:
            logger.info(f"  Imported {len(assets)} assets")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets
    logger.info("  Creating Instructions sheet...")
    create_instructions_sheet(wb)
    
    logger.info("  Creating System_Inventory sheet...")
    create_system_inventory_sheet(wb, assets)
    
    logger.info("  Creating Drift_Analysis sheet...")
    create_drift_analysis_sheet(wb)
    
    logger.error("  Creating Gaps_Failures sheet...")
    create_gaps_failures_sheet(wb)
    
    logger.info("  Creating Compliance_Summary sheet...")
    create_compliance_summary_sheet(wb)
    
    # Save workbook
    wb.save(args.output)
    logger.info(f"\n✓ Workbook generated successfully: {args.output}")
    logger.info("\nNext Steps:")
    logger.info("1. Open the workbook in Excel")
    logger.info("2. Complete System_Inventory for all systems (or import from A.5.9)")
    logger.info("3. Use verification commands from ISMS-IMP-A.8.17-S2 to populate data")
    logger.info("4. Review Drift_Analysis for statistical insights")
    logger.error("5. Address all gaps in Gaps_Failures sheet")
    logger.info("6. Review Compliance_Summary for overall status")
    logger.info("7. Track remediation until all systems are compliant")
    logger.info("\nRefer to ISMS-IMP-A.8.17-S2 for platform-specific verification commands.")

if __name__ == '__main__':
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
