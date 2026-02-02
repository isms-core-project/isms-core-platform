#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.8.4.1 - Repository Access Control Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.4: Access to Source Code
Assessment Domain 1 of 3: Repository Access Control

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific source code repositories, access control model,
and development practices.

Key customization areas:
1. Repository platforms (GitHub, GitLab, Bitbucket per your tooling)
2. Access roles and permissions (aligned with your RBAC model)
3. Team structure and ownership (based on your organization)
4. Compliance requirements (SOX, PCI-DSS per your obligations)
5. Approval workflows (per your change management)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for source code)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

**Purpose:**
Enables systematic assessment of repository access controls supporting
ISO 27001:2022 Control A.8.4 requirements for protecting source code access.

**Assessment Scope:**
- Repository inventory across all platforms
- User access matrix and permission levels
- Access approval and provisioning records
- Orphaned and stale access identification
- Service account and bot access tracking
- Access review and recertification status
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance
2. Repository Inventory - Complete repository catalog
3. User Access Matrix - Permission assignments
4. Access Approval Records - Provisioning audit trail
5. Stale Access Review - Orphaned access identification
6. Service Accounts - Bot and automation access
7. Access Recertification - Review completion status
8. Gap Analysis - Access control deficiencies
9. Evidence Register - Audit evidence tracking
10. Approval & Sign-Off - Stakeholder approval

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a84_1_repository_access.py

Requirements:
    - Python 3.8+
    - openpyxl library: pip install openpyxl

Output:
    ISMS-A84-1-Repository-Access-Assessment.xlsx
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)







# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.8.4.1"
WORKBOOK_NAME = "Repository Access Control Assessment"
CONTROL_ID = "A.8.4"
CONTROL_NAME = "Access to Source Code"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # f"{CHECK}" Green checkmark
XMARK = '\u274C'      # f"{XMARK}" Red X
WARNING = '\u26A0'    # f"{WARNING}"  Warning sign
HOURGLASS = '\u23F3'  # f"{HOURGLASS}" Hourglass
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow
QUESTION = '\u2753'   # f"{QUESTION}" Question mark
CALENDAR = '\u1F4C5' # f"{CALENDAR}" Calendar
CHECKMARK = '\u2714'  # f"{CHECKMARK}"  Check mark
MINUS = '\u2796'      # f"{MINUS}" Minus
CHART = '\U0001F4CA' # f"{CHART}" Chart
TARGET = '\U0001F3AF' # f"{TARGET}" Target

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches ISMS-IMP-A.8.4.1 specification (12 sheets)
    sheets = [
        "Instructions & Legend",
        "Repository_Inventory",
        "User_Access_Matrix",
        "Access_Approval_Records",
        "Access_Review_Log",
        "Deprovisioning_Log",
        "Third_Party_Access",
        "Service_Accounts",
        "Compliance_Scoring",
        "Gap_Analysis",
        "Evidence_Register",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    
    Returns style TEMPLATES (dictionaries), not reusable objects
    to avoid Excel repair warnings from shared Border/Font/Fill objects.
    """
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "data_cell": {
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_appropriate": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        },
        "status_excessive": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        },
        "status_orphaned": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "status_expired": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        },
        "risk_critical": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        },
        "risk_high": {
            "fill": PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),
        },
        "risk_medium": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "risk_low": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
    }
    return styles


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell. Creates NEW style objects to avoid shared object warnings."""
    if "font" in style_dict:
        cell.font = Font(
            name=style_dict["font"].name,
            size=style_dict["font"].size,
            bold=style_dict["font"].bold,
            color=style_dict["font"].color if hasattr(style_dict["font"], 'color') else None
        )
    if "fill" in style_dict:
        cell.fill = PatternFill(
            start_color=style_dict["fill"].start_color.rgb if hasattr(style_dict["fill"].start_color, 'rgb') else style_dict["fill"].start_color,
            end_color=style_dict["fill"].end_color.rgb if hasattr(style_dict["fill"].end_color, 'rgb') else style_dict["fill"].end_color,
            fill_type=style_dict["fill"].fill_type
        )
    if "alignment" in style_dict:
        cell.alignment = Alignment(
            horizontal=style_dict["alignment"].horizontal,
            vertical=style_dict["alignment"].vertical,
            wrap_text=style_dict["alignment"].wrap_text
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: DATA VALIDATIONS & COLUMN DEFINITIONS
# ============================================================================

def create_base_validations(ws):
    """Create data validation objects for standard dropdowns."""
    
    # Repository Platform validation
    dv_platform = DataValidation(
        type="list",
        formula1='"GitHub,GitLab,Bitbucket,Azure DevOps,Self-Hosted Git,Perforce,SVN,Other"',
        allow_blank=False
    )
    
    # Repository Classification validation
    dv_classification = DataValidation(
        type="list",
        formula1='"Production Code,Internal Tools,Open Source,Archived/Deprecated"',
        allow_blank=False
    )
    
    # Status validation
    dv_status = DataValidation(
        type="list",
        formula1='"Active,Archived,Deprecated,Planned"',
        allow_blank=False
    )
    
    # Criticality validation
    dv_criticality = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    
    # Yes/No validation (with emojis)
    dv_yesno = DataValidation(
        type="list",
        formula1=f'"{CHECK} Yes,❌ No,⚠️ Partial,❓ Unknown"',
        allow_blank=False
    )
    
    # Access Level validation
    dv_access_level = DataValidation(
        type="list",
        formula1='"Read,Write/Contribute,Admin"',
        allow_blank=False
    )
    
    # User Role validation
    dv_user_role = DataValidation(
        type="list",
        formula1='"Developer,Security Team,Auditor,External Contractor,Repository Admin,Service Account"',
        allow_blank=False
    )
    
    # Employment Type validation
    dv_employment = DataValidation(
        type="list",
        formula1='"Employee,Contractor,Auditor,Automated System"',
        allow_blank=False
    )
    
    # Access Status validation (with emojis)
    dv_access_status = DataValidation(
        type="list",
        formula1=f'"{CHECK} Appropriate,⚠️ Excessive,❌ Orphaned,📅 Expired,✔️ Remediated"',
        allow_blank=False
    )
    
    # Approval Status validation
    dv_approval = DataValidation(
        type="list",
        formula1=f'"{CHECK} Approved,❌ Denied,⏳ Pending"',
        allow_blank=False
    )
    
    # Compliance Status validation
    dv_compliance = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,⚠️ Partial,❌ Non-Compliant"',
        allow_blank=False
    )
    
    # Risk Level validation
    dv_risk = DataValidation(
        type="list",
        formula1='"🔴 Critical,🟠 High,🟡 Medium,🟢 Low"',
        allow_blank=False
    )
    
    # Gap Status validation
    dv_gap_status = DataValidation(
        type="list",
        formula1='"🔴 Open,🟡 In Progress,🟢 Completed,⚪ Deferred"',
        allow_blank=False
    )
    
    return {
        'platform': dv_platform,
        'classification': dv_classification,
        'status': dv_status,
        'criticality': dv_criticality,
        'yesno': dv_yesno,
        'access_level': dv_access_level,
        'user_role': dv_user_role,
        'employment': dv_employment,
        'access_status': dv_access_status,
        'approval': dv_approval,
        'compliance': dv_compliance,
        'risk': dv_risk,
        'gap_status': dv_gap_status,
    }


# ============================================================================
# SECTION 3: SHEET 1 - INSTRUCTIONS & LEGEND
# ============================================================================

def create_instructions_sheet(wb, styles):
    """Create the Instructions & Legend sheet."""
    ws = wb["Instructions & Legend"]
    
    # Header
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "ISMS-IMP-A.8.4.1 – Repository Access Control Assessment"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:F2')
    cell = ws['A2']
    cell.value = "ISO/IEC 27001:2022 - Control A.8.4: Access to Source Code"
    apply_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 25
    
    # Document Information
    row = 4
    ws[f'A{row}'] = "Document ID:"
    ws[f'B{row}'] = "ISMS-IMP-A.8.4.1"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Assessment Area:"
    ws[f'B{row}'] = "Repository Access Control"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Related Policy:"
    ws[f'B{row}'] = "ISMS-POL-A.8.4-S2"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Version:"
    ws[f'B{row}'] = "1.0"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Assessment Date:"
    ws[f'B{row}'] = "[ENTER DATE]"
    ws[f'A{row}'].font = Font(bold=True)
    apply_style(ws[f'B{row}'], styles['input_cell'])
    
    row += 1
    ws[f'A{row}'] = "Completed By:"
    ws[f'B{row}'] = "[ENTER NAME]"
    ws[f'A{row}'].font = Font(bold=True)
    apply_style(ws[f'B{row}'], styles['input_cell'])
    
    row += 1
    ws[f'A{row}'] = "Organization:"
    ws[f'B{row}'] = "[ENTER ORGANIZATION NAME]"
    ws[f'A{row}'].font = Font(bold=True)
    apply_style(ws[f'B{row}'], styles['input_cell'])
    
    row += 1
    ws[f'A{row}'] = "Review Cycle:"
    ws[f'B{row}'] = "Quarterly"
    ws[f'A{row}'].font = Font(bold=True)
    
    # How to Use This Workbook
    row += 3
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "HOW TO USE THIS WORKBOOK"
    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    
    row += 1
    instructions = [
        "1. Complete the Repository_Inventory sheet for ALL source code repositories",
        "2. Fill in the User_Access_Matrix sheet showing who has access to what",
        "3. Document all access requests and approvals in Access_Approval_Records",
        "4. Track quarterly access reviews in Access_Review_Log",
        "5. The Compliance_Scoring sheet automatically calculates compliance metrics",
        "6. Document gaps in Gap_Analysis with remediation plans",
        "7. Maintain Evidence_Register for audit traceability",
        "8. Obtain final approval and sign-off",
    ]
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Status Legend
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "STATUS LEGEND"
    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    
    row += 1
    # Legend table headers
    ws[f'A{row}'] = "Symbol"
    ws[f'B{row}'] = "Status"
    ws[f'C{row}'] = "Description"
    ws[f'D{row}'] = "Color Code"
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}{row}']
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    row += 1
    legend_data = [
        ("{CHECK}", "Appropriate", "Access is justified and appropriate", "Green"),
        ("{WARNING}", "Excessive", "User has higher access than needed", "Yellow"),
        ("{XMARK}", "Orphaned", "Former employee still has access", "Red"),
        (f"{CALENDAR}", "Expired", "Contractor access past contract end", "Red"),
        (f"{CHECKMARK}", "Remediated", "Finding has been addressed", "Blue"),
    ]
    for symbol, status, description, color in legend_data:
        ws[f'A{row}'] = symbol
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        ws[f'B{row}'] = status
        ws[f'C{row}'] = description
        ws[f'D{row}'] = color
        
        if color == "Green":
            ws[f'D{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif color == "Yellow":
            ws[f'D{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif color == "Red":
            ws[f'D{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif color == "Blue":
            ws[f'D{row}'].fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        
        row += 1
    
    # Acceptable Evidence
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "ACCEPTABLE EVIDENCE (Examples)"
    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    
    row += 1
    evidence_items = [
        "✓ Repository platform access reports (exports from GitHub, GitLab, etc.)",
        "✓ Access request tickets and approval emails",
        "✓ NDA signature records for personnel with code access",
        "✓ Contractor contracts showing access periods",
        "✓ Access review completion reports",
        "✓ Deprovisioning logs showing timely access removal",
        "✓ User-to-repository access matrix",
        "✓ HR termination records correlated with access removal",
    ]
    for evidence in evidence_items:
        ws[f'A{row}'] = evidence
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15


# ============================================================================
# SECTION 4: SHEET 2 - REPOSITORY INVENTORY
# ============================================================================

def create_repository_inventory_sheet(wb, styles, validations):
    """Create the Repository_Inventory sheet."""
    ws = wb["Repository_Inventory"]
    
    # Header
    ws.merge_cells('A1:T1')
    cell = ws['A1']
    cell.value = "REPOSITORY INVENTORY"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:T2')
    cell = ws['A2']
    cell.value = "Document all source code repositories regardless of platform"
    apply_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        "Repository Name",
        "Repository URL",
        "Repository Platform",
        "Repository Classification",
        "Repository Owner",
        "Repository Owner Email",
        "Business Purpose",
        "Primary Programming Language",
        "Last Commit Date",
        "Active/Inactive Status",
        "Total Contributors",
        "Total Commits (Lifetime)",
        "Repository Creation Date",
        "Criticality",
        "Contains Production Code",
        "Contains Sensitive Data",
        "Branch Protection Enabled",
        "Secret Scanning Enabled",
        "Last Access Review Date",
        "Notes",
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Set column widths
    widths = [25, 40, 18, 23, 20, 30, 40, 25, 18, 18, 15, 20, 20, 12, 20, 20, 23, 20, 20, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Add data validations
    ws.add_data_validation(validations['platform'])
    validations['platform'].add(f'C5:C500')
    
    ws.add_data_validation(validations['classification'])
    validations['classification'].add(f'D5:D500')
    
    ws.add_data_validation(validations['status'])
    validations['status'].add(f'J5:J500')
    
    ws.add_data_validation(validations['criticality'])
    validations['criticality'].add(f'N5:N500')
    
    ws.add_data_validation(validations['yesno'])
    validations['yesno'].add(f'O5:O500')
    validations['yesno'].add(f'P5:P500')
    validations['yesno'].add(f'Q5:Q500')
    validations['yesno'].add(f'R5:R500')
    
    # Add sample data row
    row = 4
    ws[f'A{row}'] = "[Example: customer-portal]"
    ws[f'B{row}'] = "[Example: https://github.com/org/customer-portal]"
    ws[f'C{row}'] = "GitHub"
    ws[f'D{row}'] = "Production Code"
    ws[f'E{row}'] = "[Example: Jane Doe]"
    ws[f'F{row}'] = "[Example: jane.doe@organization.com]"
    ws[f'G{row}'] = "[Example: Customer-facing web application]"
    ws[f'H{row}'] = "[Example: Python]"
    ws[f'I{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'J{row}'] = "Active"
    ws[f'K{row}'] = "[Example: 15]"
    ws[f'L{row}'] = "[Example: 3,245]"
    ws[f'M{row}'] = "[Example: 2021-03-15]"
    ws[f'N{row}'] = "Critical"
    ws[f'O{row}'] = f"{CHECK} Yes"
    ws[f'P{row}'] = f"{XMARK} No"
    ws[f'Q{row}'] = f"{CHECK} Yes"
    ws[f'R{row}'] = f"{CHECK} Yes"
    ws[f'S{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'T{row}'] = "[Example notes]"
    
    # Style sample row
    for col in range(1, 21):
        apply_style(ws.cell(row=row, column=col), styles['input_cell'])


# ============================================================================
# SECTION 5: SHEET 3 - USER ACCESS MATRIX
# ============================================================================

def create_user_access_matrix_sheet(wb, styles, validations):
    """Create the User_Access_Matrix sheet."""
    ws = wb["User_Access_Matrix"]
    
    # Header
    ws.merge_cells('A1:T1')
    cell = ws['A1']
    cell.value = "USER ACCESS MATRIX"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:T2')
    cell = ws['A2']
    cell.value = "Document all user access to repositories"
    apply_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        "User Name",
        "User Email",
        "User Role",
        "Employment Type",
        "Department/Team",
        "Repository Name",
        "Repository Platform",
        "Access Level",
        "Access Granted Date",
        "Access Approved By",
        "Business Justification",
        "Contract End Date",
        "Access Expiration Date",
        "Last Access Date",
        "Access Status",
        "Last Review Date",
        "Reviewer Comments",
        "Remediation Required",
        "Remediation Due Date",
        "Notes",
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Set column widths
    widths = [20, 30, 20, 18, 20, 25, 18, 18, 18, 20, 40, 18, 18, 18, 18, 18, 30, 18, 18, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Add data validations
    ws.add_data_validation(validations['user_role'])
    validations['user_role'].add(f'C5:C500')
    
    ws.add_data_validation(validations['employment'])
    validations['employment'].add(f'D5:D500')
    
    ws.add_data_validation(validations['platform'])
    validations['platform'].add(f'G5:G500')
    
    ws.add_data_validation(validations['access_level'])
    validations['access_level'].add(f'H5:H500')
    
    ws.add_data_validation(validations['access_status'])
    validations['access_status'].add(f'O5:O500')
    
    ws.add_data_validation(validations['yesno'])
    validations['yesno'].add(f'R5:R500')
    
    # Add sample data row
    row = 4
    ws[f'A{row}'] = "[Example: John Smith]"
    ws[f'B{row}'] = "[Example: john.smith@organization.com]"
    ws[f'C{row}'] = "Developer"
    ws[f'D{row}'] = "Employee"
    ws[f'E{row}'] = "[Example: Engineering]"
    ws[f'F{row}'] = "[Example: customer-portal]"
    ws[f'G{row}'] = "GitHub"
    ws[f'H{row}'] = "Write/Contribute"
    ws[f'I{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'J{row}'] = "[Example: Jane Doe]"
    ws[f'K{row}'] = "[Example: Active development on customer portal project]"
    ws[f'L{row}'] = ""  # N/A for employees
    ws[f'M{row}'] = ""  # N/A for employees
    ws[f'N{row}'] = (datetime.now() - timedelta(days=15)).strftime("%d.%m.%Y")
    ws[f'O{row}'] = f"{CHECK} Appropriate"
    ws[f'P{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'Q{row}'] = "[Example: Access reviewed and confirmed]"
    ws[f'R{row}'] = f"{XMARK} No"
    ws[f'S{row}'] = ""
    ws[f'T{row}'] = ""
    
    # Style sample row
    for col in range(1, 21):
        apply_style(ws.cell(row=row, column=col), styles['input_cell'])


# ============================================================================
# SECTION 6: SHEET 4 - ACCESS APPROVAL RECORDS
# ============================================================================

def create_access_approval_records_sheet(wb, styles, validations):
    """Create the Access_Approval_Records sheet."""
    ws = wb["Access_Approval_Records"]
    
    # Header
    ws.merge_cells('A1:R1')
    cell = ws['A1']
    cell.value = "ACCESS APPROVAL RECORDS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:R2')
    cell = ws['A2']
    cell.value = "Document all access request approvals for audit trail"
    apply_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        "Request Date",
        "Request ID",
        "Requestor Name",
        "Requestor Email",
        "Repository Name",
        "Access Level Requested",
        "Business Justification",
        "Expected Duration",
        "Expected End Date",
        "Repository Owner",
        "Owner Approval",
        "Approval Date",
        "Additional Approver",
        "Additional Approval",
        "Provisioning Date",
        "Provisioned By",
        "Denial Reason",
        "Notes",
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Set column widths
    widths = [15, 15, 20, 30, 25, 20, 40, 15, 18, 20, 15, 15, 20, 18, 15, 20, 30, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Add data validations
    ws.add_data_validation(validations['access_level'])
    validations['access_level'].add(f'F5:F500')
    
    dv_duration = DataValidation(
        type="list",
        formula1='"Permanent,Time-Bound"',
        allow_blank=False
    )
    ws.add_data_validation(dv_duration)
    dv_duration.add(f'H5:H500')
    
    ws.add_data_validation(validations['approval'])
    validations['approval'].add(f'K5:K500')
    
    dv_additional_approval = DataValidation(
        type="list",
        formula1=f'"{CHECK} Approved,❌ Denied,➖ N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_additional_approval)
    dv_additional_approval.add(f'N5:N500')
    
    # Add sample data row
    row = 4
    ws[f'A{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'B{row}'] = "[Example: TICKET-12345]"
    ws[f'C{row}'] = "[Example: Bob Johnson]"
    ws[f'D{row}'] = "[Example: bob.johnson@organization.com]"
    ws[f'E{row}'] = "[Example: customer-portal]"
    ws[f'F{row}'] = "Write/Contribute"
    ws[f'G{row}'] = "[Example: Need to fix authentication bug]"
    ws[f'H{row}'] = "Permanent"
    ws[f'I{row}'] = ""
    ws[f'J{row}'] = "[Example: Jane Doe]"
    ws[f'K{row}'] = f"{CHECK} Approved"
    ws[f'L{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'M{row}'] = ""
    ws[f'N{row}'] = f"{MINUS} N/A"
    ws[f'O{row}'] = (datetime.now() + timedelta(days=1)).strftime("%d.%m.%Y")
    ws[f'P{row}'] = "[Example: DevOps Team]"
    ws[f'Q{row}'] = ""
    ws[f'R{row}'] = ""
    
    # Style sample row
    for col in range(1, 19):
        apply_style(ws.cell(row=row, column=col), styles['input_cell'])


# ============================================================================
# SECTION 7: SHEET 5 - ACCESS REVIEW LOG
# ============================================================================

def create_access_review_log_sheet(wb, styles, validations):
    """Create the Access_Review_Log sheet."""
    ws = wb["Access_Review_Log"]
    
    # Header
    ws.merge_cells('A1:T1')
    cell = ws['A1']
    cell.value = "ACCESS REVIEW LOG"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:T2')
    cell = ws['A2']
    cell.value = "Track quarterly access reviews for all repositories"
    apply_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        "Review Date",
        "Review Quarter",
        "Reviewer Name",
        "Reviewer Role",
        "Repositories Reviewed",
        "Repository Names",
        "Total Users Reviewed",
        "Appropriate Access Found",
        "Excessive Access Found",
        "Unnecessary Access Found",
        "Orphaned Accounts Found",
        "Contractor Expired Found",
        "Total Findings",
        "Remediation Completed",
        "Remediation Pending",
        "Review Completion Date",
        "Days to Complete",
        "Next Review Due Date",
        "Compliance Status",
        "Notes",
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Set column widths
    widths = [15, 15, 20, 20, 15, 40, 15, 18, 18, 18, 18, 18, 15, 18, 18, 18, 15, 18, 18, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Add data validations
    ws.add_data_validation(validations['compliance'])
    validations['compliance'].add(f'S5:S500')
    
    # Add sample data row with formulas
    row = 4
    ws[f'A{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'B{row}'] = f"Q{(datetime.now().month - 1) // 3 + 1} {datetime.now().year}"
    ws[f'C{row}'] = "[Example: Repository Owner]"
    ws[f'D{row}'] = "[Example: Engineering Manager]"
    ws[f'E{row}'] = "[Example: 5]"
    ws[f'F{row}'] = "[Example: customer-portal, api-gateway, ...]"
    ws[f'G{row}'] = "[Example: 25]"
    ws[f'H{row}'] = "[Example: 22]"
    ws[f'I{row}'] = "[Example: 2]"
    ws[f'J{row}'] = "[Example: 1]"
    ws[f'K{row}'] = "[Example: 0]"
    ws[f'L{row}'] = "[Example: 0]"
    ws[f'M{row}'] = "=I4+J4+K4+L4"  # Formula: Total Findings
    ws[f'N{row}'] = "[Example: 3]"
    ws[f'O{row}'] = "=M4-N4"  # Formula: Pending = Total - Completed
    ws[f'P{row}'] = (datetime.now() + timedelta(days=7)).strftime("%d.%m.%Y")
    ws[f'Q{row}'] = "=P4-A4"  # Formula: Days to complete
    ws[f'R{row}'] = (datetime.now() + timedelta(days=90)).strftime("%d.%m.%Y")
    ws[f'S{row}'] = f"{CHECK} Compliant"
    ws[f'T{row}'] = "[Example: All findings being addressed]"
    
    # Style sample row
    for col in range(1, 21):
        apply_style(ws.cell(row=row, column=col), styles['input_cell'])


# ============================================================================
# SECTION 8: SHEET 6 - DEPROVISIONING LOG
# ============================================================================

def create_deprovisioning_log_sheet(wb, styles, validations):
    """Create the Deprovisioning_Log sheet."""
    ws = wb["Deprovisioning_Log"]
    
    # Header
    ws.merge_cells('A1:P1')
    cell = ws['A1']
    cell.value = "DEPROVISIONING LOG"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:P2')
    cell = ws['A2']
    cell.value = "Track access removal for audit trail"
    apply_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        "User Name",
        "User Email",
        "Deprovisioning Event",
        "Event Date",
        "Notification Received Date",
        "Repository Name",
        "Previous Access Level",
        "Access Removed Date",
        "Hours to Deprovision",
        "Removed By",
        "Removal Method",
        "Verification Method",
        "Verification Date",
        "Compliant Timeline",
        "Delay Reason",
        "Notes",
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Set column widths
    widths = [20, 30, 20, 15, 20, 25, 18, 18, 15, 20, 15, 18, 15, 18, 30, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Add data validations
    dv_event = DataValidation(
        type="list",
        formula1='"Termination,Role Change,Contract End,Access Review,Security Incident"',
        allow_blank=False
    )
    ws.add_data_validation(dv_event)
    dv_event.add(f'C5:C500')
    
    ws.add_data_validation(validations['access_level'])
    validations['access_level'].add(f'G5:G500')
    
    dv_method = DataValidation(
        type="list",
        formula1='"Automated,Manual"',
        allow_blank=False
    )
    ws.add_data_validation(dv_method)
    dv_method.add(f'K5:K500')
    
    dv_verification = DataValidation(
        type="list",
        formula1='"Platform Check,Login Test,API Query"',
        allow_blank=False
    )
    ws.add_data_validation(dv_verification)
    dv_verification.add(f'L5:L500')
    
    ws.add_data_validation(validations['yesno'])
    validations['yesno'].add(f'N5:N500')
    
    # Add sample data row with formulas
    row = 4
    ws[f'A{row}'] = "[Example: Former Employee]"
    ws[f'B{row}'] = "[Example: former@organization.com]"
    ws[f'C{row}'] = "Termination"
    ws[f'D{row}'] = (datetime.now() - timedelta(days=2)).strftime("%d.%m.%Y")
    ws[f'E{row}'] = (datetime.now() - timedelta(days=2)).strftime("%d.%m.%Y")
    ws[f'F{row}'] = "[Example: customer-portal]"
    ws[f'G{row}'] = "Write/Contribute"
    ws[f'H{row}'] = (datetime.now() - timedelta(days=1)).strftime("%d.%m.%Y")
    ws[f'I{row}'] = "=INT((H4-D4)*24)"  # Formula: Hours to deprovision
    ws[f'J{row}'] = "[Example: DevOps Team]"
    ws[f'K{row}'] = "Automated"
    ws[f'L{row}'] = "Platform Check"
    ws[f'M{row}'] = (datetime.now() - timedelta(days=1)).strftime("%d.%m.%Y")
    ws[f'N{row}'] = f"{CHECK} Yes"
    ws[f'O{row}'] = ""
    ws[f'P{row}'] = ""
    
    # Style sample row
    for col in range(1, 17):
        apply_style(ws.cell(row=row, column=col), styles['input_cell'])


# ============================================================================
# SECTION 9: SHEET 7 - COMPLIANCE SCORING
# ============================================================================

def create_compliance_scoring_sheet(wb, styles):
    """Create the Compliance_Scoring sheet with formulas."""
    ws = wb["Compliance_Scoring"]
    
    # Header
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "COMPLIANCE SCORING DASHBOARD"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:E2')
    cell = ws['A2']
    cell.value = "Automated compliance metrics based on assessment data"
    apply_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 25
    
    # Metrics table
    row = 4
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "COMPLIANCE METRICS"
    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    headers = ["Metric", "Current Score", "Target", "Status", "Weight"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Metrics with formulas
    row += 1
    metrics = [
        ("Repository Inventory Compliance",
         "=COUNTIF(Repository_Inventory!A:A,\"<>\")/COUNTIF(Repository_Inventory!A:A,\"<>\")*100",
         "100%", "", "15%"),
        ("Access Control Compliance",
         f'=COUNTIFS(Repository_Inventory!Q:Q,"{CHECK} Yes")/COUNTA(Repository_Inventory!Q:Q)*100',
         "100%", "", "20%"),
        ("Appropriate Access Rate",
         f'=COUNTIFS(User_Access_Matrix!O:O,"{CHECK} Appropriate")/COUNTA(User_Access_Matrix!O:O)*100',
         "≥95%", "", "25%"),
        ("Orphaned Account Rate",
         f'=COUNTIFS(User_Access_Matrix!O:O,"{XMARK} Orphaned")/COUNTA(User_Access_Matrix!O:O)*100',
         "0%", "", "15%"),
        ("Access Review Completion",
         f'=COUNTIFS(Access_Review_Log!S:S,"{CHECK} Compliant")/COUNTA(Access_Review_Log!S:S)*100',
         "100%", "", "15%"),
        ("Deprovisioning SLA Compliance",
         f'=COUNTIFS(Deprovisioning_Log!N:N,"{CHECK} Yes")/COUNTA(Deprovisioning_Log!N:N)*100',
         "≥95%", "", "10%"),
    ]
    
    for metric, formula, target, status, weight in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = formula
        ws[f'C{row}'] = target
        ws[f'D{row}'] = ""  # Status (conditional formatting will handle)
        ws[f'E{row}'] = weight
        
        # Number format for percentage
        ws[f'B{row}'].number_format = '0.0"%"'
        
        row += 1
    
    # Overall score
    row += 1
    ws[f'A{row}'] = "OVERALL REPOSITORY ACCESS COMPLIANCE SCORE"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'] = "=B6*0.15+B7*0.20+B8*0.25+(100-B9)*0.15+B10*0.15+B11*0.10"
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].number_format = '0.0"%"'
    ws[f'C{row}'] = "≥85%"
    ws[f'C{row}'].font = Font(bold=True)
    
    # Risk categorization
    row += 3
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "RISK CATEGORIZATION"
    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    risk_data = [
        ("🟢 Low Risk", "Score ≥85%", "Maintain current controls"),
        ("🟡 Medium Risk", "Score 70-84%", "Remediation within 30 days"),
        ("🔴 High Risk", "Score <70%", "Immediate remediation required"),
    ]
    
    for risk, criteria, action in risk_data:
        ws[f'A{row}'] = risk
        ws[f'B{row}'] = criteria
        ws[f'C{row}'] = action
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12


# ============================================================================
# SECTION 10: SHEET 8 - GAP ANALYSIS
# ============================================================================

def create_gap_analysis_sheet(wb, styles, validations):
    """Create the Gap_Analysis sheet."""
    ws = wb["Gap_Analysis"]
    
    # Header
    ws.merge_cells('A1:S1')
    cell = ws['A1']
    cell.value = "GAP ANALYSIS"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:S2')
    cell = ws['A2']
    cell.value = "Document identified gaps and remediation plans"
    apply_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        "Gap ID",
        "Gap Category",
        "Gap Description",
        "Policy Requirement",
        "Current State",
        "Desired State",
        "Risk Level",
        "Impact",
        "Affected Repositories",
        "Root Cause",
        "Remediation Plan",
        "Responsible Party",
        "Target Completion Date",
        "Estimated Effort",
        "Status",
        "Actual Completion Date",
        "Verification Method",
        "Verification Date",
        "Notes",
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Set column widths
    widths = [12, 18, 40, 25, 30, 30, 15, 30, 30, 30, 40, 20, 18, 18, 15, 18, 25, 15, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Add data validations
    dv_category = DataValidation(
        type="list",
        formula1='"Access Control,Inventory,Reviews,Deprovisioning,Documentation"',
        allow_blank=False
    )
    ws.add_data_validation(dv_category)
    dv_category.add(f'B5:B500')
    
    ws.add_data_validation(validations['risk'])
    validations['risk'].add(f'G5:G500')
    
    dv_effort = DataValidation(
        type="list",
        formula1='"1-2 hours,1 day,1 week,2-4 weeks,>1 month"',
        allow_blank=False
    )
    ws.add_data_validation(dv_effort)
    dv_effort.add(f'N5:N500')
    
    ws.add_data_validation(validations['gap_status'])
    validations['gap_status'].add(f'O5:O500')
    
    # Add sample data row
    row = 4
    ws[f'A{row}'] = "GAP-001"
    ws[f'B{row}'] = "Access Control"
    ws[f'C{row}'] = "[Example: Branch protection not enabled on all production repositories]"
    ws[f'D{row}'] = "[Example: REQ-2.4.1-001 - Main branch protection]"
    ws[f'E{row}'] = "[Example: 15 out of 20 production repos lack branch protection]"
    ws[f'F{row}'] = "[Example: All production repos have branch protection enabled]"
    ws[f'G{row}'] = "🟠 High"
    ws[f'H{row}'] = "[Example: Unauthorized direct commits to production code possible]"
    ws[f'I{row}'] = "[Example: api-gateway, payment-processor, ...]"
    ws[f'J{row}'] = "[Example: Repositories created before policy implementation]"
    ws[f'K{row}'] = "[Example: Enable branch protection on all production repositories]"
    ws[f'L{row}'] = "[Example: Platform Team Lead]"
    ws[f'M{row}'] = (datetime.now() + timedelta(days=30)).strftime("%d.%m.%Y")
    ws[f'N{row}'] = "1 week"
    ws[f'O{row}'] = "🟡 In Progress"
    ws[f'P{row}'] = ""
    ws[f'Q{row}'] = ""
    ws[f'R{row}'] = ""
    ws[f'S{row}'] = ""
    
    # Style sample row
    for col in range(1, 20):
        apply_style(ws.cell(row=row, column=col), styles['input_cell'])


# ============================================================================
# SECTION 11: SHEET 9 - EVIDENCE REGISTER
# ============================================================================

def create_evidence_register_sheet(wb, styles):
    """Create the Evidence_Register sheet."""
    ws = wb["Evidence_Register"]
    
    # Header
    ws.merge_cells('A1:P1')
    cell = ws['A1']
    cell.value = "EVIDENCE REGISTER"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:P2')
    cell = ws['A2']
    cell.value = "Track evidence collected for audit purposes"
    apply_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = [
        "Evidence ID",
        "Evidence Type",
        "Evidence Description",
        "Related Requirement",
        "Evidence Date",
        "Evidence Source",
        "File Name",
        "File Location",
        "Collected By",
        "Collection Date",
        "Evidence Format",
        "Retention Period",
        "Retention End Date",
        "Auditor Reviewed",
        "Auditor Comments",
        "Notes",
    ]
    
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_style(cell, styles['column_header'])
    
    # Set column widths
    widths = [12, 20, 40, 30, 15, 25, 30, 30, 20, 15, 15, 18, 18, 15, 30, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Add data validations
    dv_evidence_type = DataValidation(
        type="list",
        formula1='"Access Report,Approval Record,Review Report,Deprovisioning Log,Configuration Screenshot,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_evidence_type)
    dv_evidence_type.add(f'B5:B500')
    
    dv_format = DataValidation(
        type="list",
        formula1='"Excel,PDF,CSV,Screenshot,Email,JSON,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_format)
    dv_format.add(f'K5:K500')
    
    dv_reviewed = DataValidation(
        type="list",
        formula1=f'"{CHECK} Yes,❌ No,⏳ Pending"',
        allow_blank=False
    )
    ws.add_data_validation(dv_reviewed)
    dv_reviewed.add(f'N5:N500')
    
    # Add sample data row
    row = 4
    ws[f'A{row}'] = "EV-001"
    ws[f'B{row}'] = "Access Report"
    ws[f'C{row}'] = "[Example: GitHub organization members export Q1 2026]"
    ws[f'D{row}'] = "[Example: REQ-2.1.1-001 - RBAC implementation]"
    ws[f'E{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'F{row}'] = "[Example: GitHub API export]"
    ws[f'G{row}'] = "[Example: github-members-2026-Q1.csv]"
    ws[f'H{row}'] = "[Example: SharePoint/Evidence/A84/]"
    ws[f'I{row}'] = "[Example: Security Analyst]"
    ws[f'J{row}'] = datetime.now().strftime("%d.%m.%Y")
    ws[f'K{row}'] = "CSV"
    ws[f'L{row}'] = "3 years"
    ws[f'M{row}'] = (datetime.now() + timedelta(days=1095)).strftime("%d.%m.%Y")
    ws[f'N{row}'] = f"{HOURGLASS} Pending"
    ws[f'O{row}'] = ""
    ws[f'P{row}'] = ""
    
    # Style sample row
    for col in range(1, 17):
        apply_style(ws.cell(row=row, column=col), styles['input_cell'])


# ============================================================================
# SECTION 12: SHEET 10 - APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff_sheet(wb, styles):
    """Create the Approval_Sign_Off sheet."""
    ws = wb["Approval_Sign_Off"]
    
    # Header
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "ASSESSMENT COMPLETION CERTIFICATION"
    apply_style(cell, styles['header'])
    ws.row_dimensions[1].height = 40
    
    # Certification text
    row = 3
    ws.merge_cells(f'A{row}:E{row+5}')
    cell = ws[f'A{row}']
    cert_text = """I certify that this repository access control assessment has been completed in accordance with ISMS-POL-A.8.4 requirements and that the information documented herein is accurate to the best of my knowledge.

Assessment Period: [Enter Quarter/Year]
Assessment Completion Date: [Enter Date]
Overall Compliance Score: [See Compliance_Scoring sheet]"""
    cell.value = cert_text
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font = Font(name="Calibri", size=11)
    ws.row_dimensions[row].height = 120
    
    # Approvals section
    row += 7
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "APPROVALS"
    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Signature blocks
    signatures = [
        ("Repository Owners", 3),
        ("Information Security Manager", 1),
        ("Chief Information Security Officer (CISO)", 1),
    ]
    
    for sig_title, count in signatures:
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = f"{sig_title}:"
        cell.font = Font(bold=True)
        
        for i in range(count):
            row += 1
            ws[f'A{row}'] = "Name:"
            ws[f'B{row}'] = "________________"
            apply_style(ws[f'B{row}'], styles['input_cell'])
            ws[f'C{row}'] = "Signature:"
            ws[f'D{row}'] = "________________"
            apply_style(ws[f'D{row}'], styles['input_cell'])
            ws[f'E{row}'] = "Date: ______"
            apply_style(ws[f'E{row}'], styles['input_cell'])
    
    # Next review
    row += 3
    ws[f'A{row}'] = "Next Scheduled Review:"
    ws[f'B{row}'] = "[Completion Date + 90 days]"
    ws[f'A{row}'].font = Font(bold=True)
    apply_style(ws[f'B{row}'], styles['input_cell'])
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15


# ============================================================================
# SECTION 13: MAIN FUNCTION
# ============================================================================

def main():
    """Main function to generate the Excel workbook."""
    logger.info("Starting ISMS-A84-1 Repository Access Assessment Workbook Generation...")
    
    # Create workbook and get styles
    wb = create_workbook()
    styles = setup_styles()
    
    # Get first sheet to create validations
    first_sheet = wb["Instructions & Legend"]
    validations = create_base_validations(first_sheet)
    
    # Create all sheets
    logger.info("Creating Instructions & Legend sheet...")
    create_instructions_sheet(wb, styles)
    
    logger.info("Creating Repository_Inventory sheet...")
    create_repository_inventory_sheet(wb, styles, validations)
    
    logger.info("Creating User_Access_Matrix sheet...")
    create_user_access_matrix_sheet(wb, styles, validations)
    
    logger.info("Creating Access_Approval_Records sheet...")
    create_access_approval_records_sheet(wb, styles, validations)
    
    logger.info("Creating Access_Review_Log sheet...")
    create_access_review_log_sheet(wb, styles, validations)
    
    logger.info("Creating Deprovisioning_Log sheet...")
    create_deprovisioning_log_sheet(wb, styles, validations)
    
    logger.info("Creating Compliance_Scoring sheet...")
    create_compliance_scoring_sheet(wb, styles)
    
    logger.info("Creating Gap_Analysis sheet...")
    create_gap_analysis_sheet(wb, styles, validations)
    
    logger.info("Creating Evidence_Register sheet...")
    create_evidence_register_sheet(wb, styles)
    
    logger.info("Creating Approval_Sign_Off sheet...")
    create_approval_signoff_sheet(wb, styles)
    
    # Save workbook
    output_filename = f"ISMS-IMP-A.8.4.1_Repository_Access_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(output_filename)
    logger.info(f"\n✅ Workbook generated successfully: {output_filename}")
    logger.info(f"{CHART} Sheets created: {len(wb.sheetnames)}")
    logger.info("\nWorkbook is ready for use in repository access control assessments!")


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
