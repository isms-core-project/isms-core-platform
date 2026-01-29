#!/usr/bin/env python3
"""
isms_check_circular_refs.py
Automated detection of circular references in Excel workbooks

Usage:
    python3 isms_check_circular_refs.py <workbook.xlsx>
    python3 isms_check_circular_refs.py *.xlsx  # Test multiple files

Returns:
    Exit code 0 if no circular references found
    Exit code 1 if circular references detected
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import re
import sys
import os

def check_circular_reference(ws, row, col, formula):
    """
    Check if a formula creates a circular reference.
    
    Args:
        ws: Worksheet object
        row: Row number (1-indexed)
        col: Column number (1-indexed)
        formula: Formula string to check
    
    Returns:
        str: Error message if circular reference found
        None: If no circular reference
    """
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return None
    
    current_cell = f"{get_column_letter(col)}{row}"
    
    # Check 1: Direct self-reference (A1 references A1)
    # Match patterns like A1, $A$1, A$1, $A1
    pattern = r'\$?[A-Z]+\$?\d+'
    refs = re.findall(pattern, formula)
    
    for ref in refs:
        # Remove $ signs for comparison
        clean_ref = ref.replace('$', '')
        if clean_ref == current_cell:
            return f"DIRECT CIRCULAR: {current_cell} references itself"
    
    # Check 2: Range inclusion (A1 is within range A1:A10)
    range_pattern = r'([A-Z]+\d+):([A-Z]+\d+)'
    ranges = re.findall(range_pattern, formula)
    
    for start, end in ranges:
        try:
            # Parse start cell
            start_match = re.match(r'([A-Z]+)(\d+)', start)
            if not start_match:
                continue
            start_col = column_index_from_string(start_match.group(1))
            start_row = int(start_match.group(2))
            
            # Parse end cell
            end_match = re.match(r'([A-Z]+)(\d+)', end)
            if not end_match:
                continue
            end_col = column_index_from_string(end_match.group(1))
            end_row = int(end_match.group(2))
            
            # Check if current cell is within this range
            if (start_col <= col <= end_col and start_row <= row <= end_row):
                return f"RANGE CIRCULAR: {current_cell} is within range {start}:{end}"
        except Exception as e:
            # Skip malformed references
            pass
    
    return None

def scan_workbook(filepath, verbose=True):
    """
    Scan entire workbook for circular references.
    
    Args:
        filepath: Path to Excel workbook
        verbose: Print detailed output
    
    Returns:
        List of issues found (empty list if clean)
    """
    
    if verbose:
        print("="*80)
        print(f"SCANNING: {os.path.basename(filepath)}")
        print("="*80)
    
    try:
        wb = load_workbook(filepath, data_only=False)
    except Exception as e:
        if verbose:
            print(f"❌ ERROR: Could not load workbook: {e}")
        return [f"Load error: {e}"]
    
    all_issues = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_issues = []
        
        # Scan reasonable range (adjust if needed for very large sheets)
        max_row = min(ws.max_row + 10, 500)  # Scan up to 500 rows
        max_col = min(ws.max_column + 5, 50)  # Scan up to 50 columns
        
        for row in range(1, max_row):
            for col in range(1, max_col):
                try:
                    cell = ws.cell(row, col)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        result = check_circular_reference(ws, row, col, cell.value)
                        if result:
                            issue = f"  {result}"
                            issue += f"\n  Formula: {cell.value}"
                            sheet_issues.append(issue)
                except Exception:
                    # Skip cells that can't be read
                    pass
        
        if sheet_issues:
            if verbose:
                print(f"\n❌ {sheet_name}:")
                for issue in sheet_issues:
                    print(issue)
            all_issues.extend(sheet_issues)
        else:
            if verbose:
                print(f"✅ {sheet_name}: Clean")
    
    if verbose:
        print("\n" + "="*80)
        if all_issues:
            print(f"❌ FOUND {len(all_issues)} CIRCULAR REFERENCE(S)")
        else:
            print("✅ NO CIRCULAR REFERENCES DETECTED")
        print("="*80)
    
    return all_issues

def main():
    """Main entry point."""
    
    if len(sys.argv) < 2:
        print("Usage: python3 isms_check_circular_refs.py <workbook.xlsx> [workbook2.xlsx ...]")
        print("\nExamples:")
        print("  python3 isms_check_circular_refs.py dashboard.xlsx")
        print("  python3 isms_check_circular_refs.py *.xlsx")
        print("  python3 isms_check_circular_refs.py ISMS-IMP-A.8.*.xlsx")
        sys.exit(1)
    
    files_to_test = sys.argv[1:]
    total_files = len(files_to_test)
    failed_files = []
    
    # If testing multiple files, show summary format
    multi_file = total_files > 1
    
    if multi_file:
        print(f"\nTesting {total_files} file(s)...\n")
    
    for filepath in files_to_test:
        if not os.path.exists(filepath):
            print(f"❌ File not found: {filepath}")
            failed_files.append(filepath)
            continue
        
        issues = scan_workbook(filepath, verbose=not multi_file)
        
        if issues:
            failed_files.append(filepath)
            if multi_file:
                print(f"❌ {os.path.basename(filepath)}: {len(issues)} issue(s)")
        else:
            if multi_file:
                print(f"✅ {os.path.basename(filepath)}: Clean")
    
    # Summary for multiple files
    if multi_file:
        print("\n" + "="*80)
        print(f"SUMMARY: {total_files - len(failed_files)}/{total_files} passed")
        
        if failed_files:
            print("\nFailed files:")
            for f in failed_files:
                print(f"  - {os.path.basename(f)}")
        print("="*80)
    
    # Exit code: 0 if all passed, 1 if any failed
    return 1 if failed_files else 0

if __name__ == "__main__":
    sys.exit(main())
