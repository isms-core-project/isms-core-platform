#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.29.1 - Security Controls During Disruption Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.29: Information Security During Disruption
Assessment Domain 1 of 4: Security Controls During Disruption

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for documenting
security controls that must be maintained during disruptive events.

**Purpose:**
Enables systematic inventory of security controls, minimum baseline definition,
security posture levels, and BC/DR security review tracking.

**Generated Workbook Structure:**
1. Instructions - Assessment guidance and methodology
2. Security_Control_Inventory - Master security control catalogue
3. Minimum_Baseline - Non-negotiable security controls
4. Security_Posture_Levels - Tiered security definitions
5. Compensating_Controls - Alternative control measures
6. BCDR_Security_Review - BC/DR plan security approval tracking
7. Evidence_Register - Audit evidence tracking
8. Approval_SignOff - Stakeholder review and approval

================================================================================
"""

# =============================================================================
# IMPORTS - Standard Library
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# IMPORTS - Third Party
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.29.1"
WORKBOOK_NAME = "Security Controls During Disruption"
CONTROL_ID = "A.5.29"
CONTROL_NAME = "Information Security During Disruption"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches specification
    sheets = [
        "Instructions",
        "Security_Control_Inventory",
        "Minimum_Baseline",
        "Security_Posture_Levels",
        "Compensating_Controls",
        "BCDR_Security_Review",
        "Evidence_Register",
        "Approval_SignOff",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet with assessment guidance."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Security Controls During Disruption"),
        ("Control Reference", CONTROL_ID),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Annual"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Complete the Security_Control_Inventory sheet with all security controls",
        "2. Define Minimum_Baseline for non-negotiable controls",
        "3. Document Security_Posture_Levels with transition criteria",
        "4. Identify Compensating_Controls for degradable controls",
        "5. Complete BCDR_Security_Review for all BC/DR plans",
        "6. Link evidence in Evidence_Register",
        "7. Obtain approvals in Approval_SignOff sheet",
    ]

    row += 1
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    row += 1
    ws[f"A{row}"] = "CONTROL CATEGORIES"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    categories = [
        ("Access Control", "Authentication, authorisation, privileged access"),
        ("Data Encryption", "TLS, disk encryption, key management"),
        ("Logging", "Audit trails, SIEM, event correlation"),
        ("Network Security", "Firewalls, IDS/IPS, segmentation"),
        ("Backup Protection", "Encrypted backups, offsite storage"),
        ("Endpoint Security", "EDR, antimalware, patching"),
        ("Physical Security", "CCTV, access badges, secure areas"),
    ]

    row += 1
    for cat, desc in categories:
        ws[f"A{row}"] = cat
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = desc
        row += 1

    row += 1
    ws[f"A{row}"] = "DISRUPTION CLASSIFICATION"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    classifications = [
        ("Non-Negotiable", "Must be maintained at all times"),
        ("Degradable", "Can be reduced with compensating controls and approval"),
        ("Deferrable", "Can be postponed temporarily"),
        ("Not Applicable", "Not relevant during disruption"),
    ]

    row += 1
    for cls, desc in classifications:
        ws[f"A{row}"] = cls
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = desc
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    ws.freeze_panes = "A4"


# =============================================================================
# SECURITY CONTROL INVENTORY SHEET
# =============================================================================
def create_security_control_inventory_sheet(ws, styles):
    """Create the Security_Control_Inventory sheet - master catalogue."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "SECURITY CONTROL INVENTORY - Master Catalogue"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Column definitions per specification
    columns = [
        ("Control_ID", 15),
        ("Control_Name", 35),
        ("Control_Category", 20),
        ("ISO_Reference", 15),
        ("Normal_Status", 15),
        ("Disruption_Classification", 22),
        ("Recovery_Priority", 15),
        ("Compensating_Control_ID", 18),
        ("Owner", 25),
        ("Last_Review_Date", 16),
        ("Notes", 40),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data validations
    dv_category = DataValidation(
        type="list",
        formula1='"Access Control,Data Encryption,Logging,Network Security,Backup Protection,Endpoint Security,Physical Security,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_category)

    dv_status = DataValidation(
        type="list",
        formula1='"Operational,Partial,Not Applicable"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_classification = DataValidation(
        type="list",
        formula1='"Non-Negotiable,Degradable,Deferrable,Not Applicable"',
        allow_blank=False
    )
    ws.add_data_validation(dv_classification)

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_priority)

    # Pre-populate with common security controls
    sample_controls = [
        ("SC-001", "Multi-Factor Authentication", "Access Control", "A.8.5", "Operational", "Degradable", "Critical"),
        ("SC-002", "Role-Based Access Control", "Access Control", "A.5.15", "Operational", "Non-Negotiable", "Critical"),
        ("SC-003", "Privileged Access Management", "Access Control", "A.8.2", "Operational", "Non-Negotiable", "Critical"),
        ("SC-004", "Data Encryption at Rest", "Data Encryption", "A.8.24", "Operational", "Non-Negotiable", "Critical"),
        ("SC-005", "TLS for Data in Transit", "Data Encryption", "A.8.24", "Operational", "Non-Negotiable", "Critical"),
        ("SC-006", "Security Event Logging", "Logging", "A.8.15", "Operational", "Non-Negotiable", "Critical"),
        ("SC-007", "SIEM Correlation", "Logging", "A.8.16", "Operational", "Degradable", "High"),
        ("SC-008", "Firewall Protection", "Network Security", "A.8.20", "Operational", "Non-Negotiable", "Critical"),
        ("SC-009", "Network Segmentation", "Network Security", "A.8.22", "Operational", "Non-Negotiable", "Critical"),
        ("SC-010", "IDS/IPS", "Network Security", "A.8.20", "Operational", "Degradable", "High"),
        ("SC-011", "Backup Encryption", "Backup Protection", "A.8.13", "Operational", "Non-Negotiable", "Critical"),
        ("SC-012", "Endpoint Detection & Response", "Endpoint Security", "A.8.7", "Operational", "Degradable", "High"),
        ("SC-013", "Antimalware Protection", "Endpoint Security", "A.8.7", "Operational", "Non-Negotiable", "Critical"),
        ("SC-014", "Patch Management", "Endpoint Security", "A.8.8", "Operational", "Deferrable", "High"),
        ("SC-015", "Physical Access Controls", "Physical Security", "A.7.1", "Operational", "Non-Negotiable", "High"),
    ]

    for row_idx, control in enumerate(sample_controls, start=4):
        for col_idx, value in enumerate(control, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]

    # Additional blank rows for input
    for r in range(len(sample_controls) + 4, len(sample_controls) + 54):
        ws.cell(row=r, column=1, value=f"SC-{r-3:03d}").font = Font(color="808080")

        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Apply validations
        dv_category.add(ws.cell(row=r, column=3))
        dv_status.add(ws.cell(row=r, column=5))
        dv_classification.add(ws.cell(row=r, column=6))
        dv_priority.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "C4"


# =============================================================================
# MINIMUM BASELINE SHEET
# =============================================================================
def create_minimum_baseline_sheet(ws, styles):
    """Create the Minimum_Baseline sheet - non-negotiable controls."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "MINIMUM SECURITY BASELINE - Non-Negotiable Controls"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Control_ID", 15),
        ("Control_Name", 35),
        ("Category", 20),
        ("Minimum_Requirement", 45),
        ("Rationale", 45),
        ("Never_Acceptable_Actions", 45),
        ("Approval_Status", 15),
        ("Approved_By", 25),
        ("Approval_Date", 16),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate with baseline requirements
    baseline_controls = [
        ("SC-002", "Role-Based Access Control", "Access Control",
         "Authentication required for all system access",
         "Prevents unauthorised access during chaos",
         "Removal of authentication requirements"),
        ("SC-004", "Data Encryption at Rest", "Data Encryption",
         "Encryption at rest for confidential+ data",
         "Data remains protected if media lost",
         "Decryption of data without re-encryption"),
        ("SC-006", "Security Event Logging", "Logging",
         "Critical system logging continues",
         "Audit trail for post-incident investigation",
         "Disabled logging on critical systems"),
        ("SC-008", "Firewall Protection", "Network Security",
         "Critical network boundaries maintained",
         "Prevents lateral movement during recovery",
         "Disabling of network security controls"),
        ("SC-011", "Backup Encryption", "Backup Protection",
         "Backups remain encrypted and access-controlled",
         "Prevents backup compromise for data theft",
         "Sharing of privileged credentials without accountability"),
    ]

    dv_status = DataValidation(
        type="list",
        formula1='"Pending,Approved,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for row_idx, control in enumerate(baseline_controls, start=4):
        for col_idx, value in enumerate(control, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            if col_idx <= 6:
                cell.alignment = Alignment(wrap_text=True, vertical="center")

        # Approval status column
        ws.cell(row=row_idx, column=7).fill = styles["input_cell"]["fill"]
        dv_status.add(ws.cell(row=row_idx, column=7))
        ws.cell(row=row_idx, column=8).fill = styles["input_cell"]["fill"]
        ws.cell(row=row_idx, column=9).fill = styles["input_cell"]["fill"]

    # Additional blank rows
    for r in range(len(baseline_controls) + 4, len(baseline_controls) + 24):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_status.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "C4"


# =============================================================================
# SECURITY POSTURE LEVELS SHEET
# =============================================================================
def create_security_posture_levels_sheet(ws, styles):
    """Create the Security_Posture_Levels sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "SECURITY POSTURE LEVELS - Tiered Security Definitions"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Posture_Level", 15),
        ("Description", 40),
        ("Trigger_Conditions", 40),
        ("Control_Status", 35),
        ("Monitoring_Enhancement", 35),
        ("Transition_To", 25),
        ("Transition_Authority", 30),
        ("Documentation_Required", 35),
        ("Example_Scenario", 40),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate with posture levels
    posture_levels = [
        ("Normal", "Full security controls operational", "No disruption",
         "All controls at 100%", "Standard monitoring",
         "Elevated", "N/A (default state)", "None",
         "Day-to-day operations"),
        ("Elevated", "Enhanced monitoring, accelerated patching", "Minor disruption detected",
         "All controls operational, enhanced vigilance", "Increased log review frequency, lowered alert thresholds",
         "Normal, Degraded", "CISO or Security Team Lead", "Incident ticket",
         "Single system failure, minor security incident"),
        ("Degraded", "Core controls maintained, non-critical deferred", "Moderate disruption",
         "Essential controls only, some relaxations", "24/7 security monitoring, external threat intel",
         "Elevated, Emergency", "CISO + CIO", "Formal notification to Executive Management",
         "Data centre failover, regional outage"),
        ("Emergency", "Minimum baseline only, survival mode", "Severe disruption",
         "Minimum baseline controls only", "Real-time monitoring of critical systems, threat hunting",
         "Degraded, Recovery", "Executive Management (CEO or delegate)", "Emergency declaration document",
         "Multiple site disaster, major cyber attack"),
        ("Recovery", "Gradual restoration with security validation", "Returning to normal",
         "Phased control restoration", "Validation monitoring, anomaly detection",
         "Degraded, Normal", "CISO approval for each phase", "Phase completion checklist",
         "Post-disaster recovery phase"),
    ]

    for row_idx, level in enumerate(posture_levels, start=4):
        for col_idx, value in enumerate(level, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.freeze_panes = "B4"


# =============================================================================
# COMPENSATING CONTROLS SHEET
# =============================================================================
def create_compensating_controls_sheet(ws, styles):
    """Create the Compensating_Controls sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "COMPENSATING CONTROLS - Alternative Measures"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Compensating_ID", 15),
        ("Primary_Control_ID", 18),
        ("Primary_Control_Name", 30),
        ("Compensating_Measure", 45),
        ("Effectiveness", 15),
        ("Implementation_Requirements", 40),
        ("Activation_Trigger", 35),
        ("Duration_Limit", 20),
        ("Test_Status", 15),
        ("Last_Test_Date", 16),
        ("Test_Results", 40),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate with sample compensating controls
    compensating = [
        ("CC-001", "SC-001", "Multi-Factor Authentication",
         "Enhanced logging, session limits, IP restrictions",
         "Partial", "SOC to enable enhanced monitoring, IT to apply session limits",
         "MFA infrastructure unavailable", "72 hours",
         "Tested", "", "Effective during Q3 drill"),
        ("CC-002", "SC-007", "SIEM Correlation",
         "Manual log review for critical systems",
         "Partial", "Security team manual review every 4 hours",
         "SIEM platform unavailable", "48 hours",
         "Tested", "", "Functional but resource-intensive"),
        ("CC-003", "SC-010", "IDS/IPS",
         "Enhanced firewall logging, manual traffic analysis",
         "Partial", "Network team to review firewall logs hourly",
         "IDS/IPS system failure", "24 hours",
         "Tested", "", "Acceptable for short duration"),
        ("CC-004", "SC-012", "EDR",
         "Increased antimalware scans, manual endpoint review",
         "Partial", "IT to run full scans twice daily",
         "EDR platform unavailable", "48 hours",
         "Untested", "", ""),
        ("CC-005", "SC-014", "Patch Management",
         "Critical patches only, manual review process",
         "Partial", "Security team manual patch review",
         "Patch management system unavailable", "30 days for non-critical",
         "Tested", "", "Used during system upgrade"),
    ]

    dv_effectiveness = DataValidation(
        type="list",
        formula1='"Full,Partial,Minimal"',
        allow_blank=False
    )
    ws.add_data_validation(dv_effectiveness)

    dv_test_status = DataValidation(
        type="list",
        formula1='"Tested,Untested,Failed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_test_status)

    for row_idx, control in enumerate(compensating, start=4):
        for col_idx, value in enumerate(control, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            if col_idx in [4, 6, 7, 11]:
                cell.alignment = Alignment(wrap_text=True, vertical="center")

        dv_effectiveness.add(ws.cell(row=row_idx, column=5))
        dv_test_status.add(ws.cell(row=row_idx, column=9))

    # Additional blank rows
    for r in range(len(compensating) + 4, len(compensating) + 24):
        ws.cell(row=r, column=1, value=f"CC-{r-3:03d}").font = Font(color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_effectiveness.add(ws.cell(row=r, column=5))
        dv_test_status.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "D4"


# =============================================================================
# BC/DR SECURITY REVIEW SHEET
# =============================================================================
def create_bcdr_security_review_sheet(ws, styles):
    """Create the BCDR_Security_Review sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "BC/DR SECURITY REVIEW - Plan Security Approval Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Plan_ID", 15),
        ("Plan_Name", 35),
        ("Plan_Type", 20),
        ("Plan_Owner", 25),
        ("Security_Section_Present", 18),
        ("CISO_Review_Date", 16),
        ("CISO_Approval_Status", 18),
        ("Gaps_Identified", 40),
        ("Remediation_Due_Date", 18),
        ("Remediation_Status", 18),
        ("Next_Review_Due", 16),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_plan_type = DataValidation(
        type="list",
        formula1='"BCP,DRP,Crisis Management,IT Recovery,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_plan_type)

    dv_security_section = DataValidation(
        type="list",
        formula1='"Yes,No,Partial"',
        allow_blank=False
    )
    ws.add_data_validation(dv_security_section)

    dv_approval = DataValidation(
        type="list",
        formula1='"Approved,Rejected,Pending"',
        allow_blank=False
    )
    ws.add_data_validation(dv_approval)

    dv_remediation = DataValidation(
        type="list",
        formula1='"Open,In Progress,Closed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_remediation)

    # Data entry rows
    for r in range(4, 34):
        ws.cell(row=r, column=1, value=f"BCP-{r-3:03d}").font = Font(color="808080")

        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_plan_type.add(ws.cell(row=r, column=3))
        dv_security_section.add(ws.cell(row=r, column=5))
        dv_approval.add(ws.cell(row=r, column=7))
        dv_remediation.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "C4"


# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================
def create_evidence_register_sheet(ws, styles):
    """Create the Evidence_Register sheet for audit traceability."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 45),
        ("Related_Section", 25),
        ("Collection_Date", 16),
        ("Location", 40),
        ("Collected_By", 25),
        ("Valid_Until", 16),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Evidence type validation
    dv_type = DataValidation(
        type="list",
        formula1='"Document,Approval,Test Result,Configuration,Screenshot,Attestation"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    # Data entry rows
    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")

        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "C4"


# =============================================================================
# APPROVAL SIGN-OFF SHEET
# =============================================================================
def create_approval_signoff_sheet(ws, styles):
    """Create the Approval_SignOff sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Total Security Controls", "=COUNTA(Security_Control_Inventory!A4:A103)-COUNTBLANK(Security_Control_Inventory!B4:B103)"),
        ("Non-Negotiable Controls", '=COUNTIF(Security_Control_Inventory!F4:F103,"Non-Negotiable")'),
        ("BC/DR Plans Reviewed", '=COUNTIF(BCDR_Security_Review!G4:G33,"Approved")'),
        ("Compensating Controls Tested", '=COUNTIF(Compensating_Controls!I4:I33,"Tested")'),
        ("Open Remediation Items", '=COUNTIF(BCDR_Security_Review!J4:J33,"Open")'),
        ("Assessment Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "" or not str(value).startswith("="):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    # Assessment Completed By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    completion_fields = ["Name", "Role/Title", "Department", "Email", "Date"]
    row += 1
    for field in completion_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Reviewed By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (Security Manager)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    review_fields = ["Name", "Date", "Signature"]
    row += 1
    for field in review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Approved By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    approval_fields = ["Name", "Date", "Approval Decision", "Signature"]
    row += 1
    for field in approval_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Decision dropdown
    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add(ws[f"B{row-2}"])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A3"


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main() -> int:
    """
    Main execution function - orchestrates workbook creation.

    Returns:
        int: 0 on success, 1 on failure
    """
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/8] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/8] Creating Security_Control_Inventory sheet...")
        create_security_control_inventory_sheet(wb["Security_Control_Inventory"], styles)

        logger.info("[3/8] Creating Minimum_Baseline sheet...")
        create_minimum_baseline_sheet(wb["Minimum_Baseline"], styles)

        logger.info("[4/8] Creating Security_Posture_Levels sheet...")
        create_security_posture_levels_sheet(wb["Security_Posture_Levels"], styles)

        logger.info("[5/8] Creating Compensating_Controls sheet...")
        create_compensating_controls_sheet(wb["Compensating_Controls"], styles)

        logger.info("[6/8] Creating BCDR_Security_Review sheet...")
        create_bcdr_security_review_sheet(wb["BCDR_Security_Review"], styles)

        logger.info("[7/8] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[8/8] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("Next steps:")
        logger.info("  1) Complete document information in Instructions")
        logger.info("  2) Inventory security controls")
        logger.info("  3) Define minimum baseline")
        logger.info("  4) Document security posture levels")
        logger.info("  5) Identify compensating controls")
        logger.info("  6) Review BC/DR plans for security")
        logger.info("  7) Link evidence")
        logger.info("  8) Obtain approvals")
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        logger.error("Install with: pip install openpyxl")
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# END OF SCRIPT
# =============================================================================
# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.29.1 specification
# =============================================================================
