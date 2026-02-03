#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.29.4 - Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.29: Information Security During Disruption
Assessment Domain 4 of 4: Compliance Dashboard

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates an Excel workbook providing executive-level visibility
into security-during-disruption compliance status.

**Generated Workbook Structure:**
1. Executive_Dashboard - High-level summary
2. BCDR_Security_Status - BC/DR plan coverage
3. Emergency_Access_Status - Break-glass readiness
4. Personnel_Status - Team availability
5. Security_Debt_Status - Debt aging analysis
6. Disruption_History - Past incident analysis
7. Trend_Analysis - Historical performance
8. Data_Sources - Source workbook references
9. Instructions - Guidance
10. Approval_SignOff - Dashboard approval

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.29.4"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.5.29"
CONTROL_NAME = "Information Security During Disruption"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "dashboard_header": {
            "font": Font(name="Calibri", size=18, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "metric_label": {
            "font": Font(name="Calibri", size=12, bold=True),
            "alignment": Alignment(horizontal="left", vertical="center"),
        },
        "metric_value": {
            "font": Font(name="Calibri", size=24, bold=True),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "green": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "yellow": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "red": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Executive_Dashboard",
        "BCDR_Security_Status",
        "Emergency_Access_Status",
        "Personnel_Status",
        "Security_Debt_Status",
        "Disruption_History",
        "Trend_Analysis",
        "Data_Sources",
        "Instructions",
        "Approval_SignOff",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# EXECUTIVE DASHBOARD SHEET
# =============================================================================
def create_executive_dashboard_sheet(ws, styles):
    """Create Executive_Dashboard sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "A.5.29 INFORMATION SECURITY DURING DISRUPTION - EXECUTIVE DASHBOARD"
    ws["A1"].font = styles["dashboard_header"]["font"]
    ws["A1"].fill = styles["dashboard_header"]["fill"]
    ws["A1"].alignment = styles["dashboard_header"]["alignment"]
    ws.row_dimensions[1].height = 45

    # Overall Compliance Score
    ws.merge_cells("A3:B3")
    ws["A3"] = "Overall Compliance Score"
    ws["A3"].font = Font(bold=True, size=14)

    ws.merge_cells("A4:B6")
    ws["A4"] = "85%"  # Placeholder - would be formula in real use
    ws["A4"].font = Font(name="Calibri", size=48, bold=True, color="228B22")
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A4"].fill = styles["green"]["fill"]

    ws.merge_cells("C3:F3")
    ws["C3"] = "Dashboard Period"
    ws["C3"].font = Font(bold=True, size=12)
    ws.merge_cells("C4:F4")
    ws["C4"] = f"Generated: {GENERATED_DATE}"

    # Key Metrics
    row = 8
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "KEY METRICS"
    ws[f"A{row}"].font = styles["header"]["font"]
    ws[f"A{row}"].fill = styles["header"]["fill"]
    ws[f"A{row}"].alignment = styles["header"]["alignment"]

    metrics = [
        ("BC/DR Plans with CISO Approval", "Target: 100%", ""),
        ("Emergency Access Tests Completed", "Target: 100% annually", ""),
        ("Security Personnel Availability", "Target: 100% reachable within 1h", ""),
        ("Security Debt Items >90 Days", "Target: 0", ""),
        ("Recovery Site Security Findings", "Target: 0 critical/high", ""),
    ]

    row += 1
    for metric, target, value in metrics:
        ws[f"A{row}"] = metric
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"C{row}"] = target
        ws[f"E{row}"] = value
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].border = styles["border"]
        row += 1

    # Current Issues
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "CURRENT ISSUES & ALERTS"
    ws[f"A{row}"].font = styles["header"]["font"]
    ws[f"A{row}"].fill = styles["header"]["fill"]
    ws[f"A{row}"].alignment = styles["header"]["alignment"]

    row += 1
    ws[f"A{row}"] = "Issue Description"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"D{row}"] = "Status"
    ws[f"D{row}"].font = Font(bold=True)
    ws[f"E{row}"] = "Owner"
    ws[f"E{row}"].font = Font(bold=True)

    for _ in range(5):
        row += 1
        ws[f"A{row}"].fill = styles["input_cell"]["fill"]
        ws[f"A{row}"].border = styles["border"]
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"D{row}"].border = styles["border"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].border = styles["border"]

    # CISO Commentary
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "CISO COMMENTARY"
    ws[f"A{row}"].font = styles["header"]["font"]
    ws[f"A{row}"].fill = styles["header"]["fill"]
    ws[f"A{row}"].alignment = styles["header"]["alignment"]

    row += 1
    ws.merge_cells(f"A{row}:F{row+4}")
    ws[f"A{row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{row}"].border = styles["border"]
    ws[f"A{row}"].alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 80

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15


# =============================================================================
# BC/DR SECURITY STATUS SHEET
# =============================================================================
def create_bcdr_security_status_sheet(ws, styles):
    """Create BCDR_Security_Status sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "BC/DR SECURITY STATUS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Plan_ID", 15),
        ("Plan_Name", 35),
        ("Plan_Type", 18),
        ("Plan_Owner", 25),
        ("Security_Section", 15),
        ("CISO_Approved", 15),
        ("Approval_Date", 16),
        ("Next_Review", 16),
        ("Review_Overdue", 15),
        ("Status", 12),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_bool = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=False)
    ws.add_data_validation(dv_bool)

    dv_approved = DataValidation(type="list", formula1='"Yes,No,Pending"', allow_blank=False)
    ws.add_data_validation(dv_approved)

    for r in range(4, 34):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        dv_bool.add(ws.cell(row=r, column=5))
        dv_approved.add(ws.cell(row=r, column=6))

        # Review overdue formula
        ws.cell(row=r, column=9).value = f'=IF(H{r}<>"",IF(H{r}<TODAY(),"Yes","No"),"")'
        ws.cell(row=r, column=9).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.freeze_panes = "C4"


# =============================================================================
# EMERGENCY ACCESS STATUS SHEET
# =============================================================================
def create_emergency_access_status_sheet(ws, styles):
    """Create Emergency_Access_Status sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "EMERGENCY ACCESS STATUS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Account_ID", 15),
        ("Account_Name", 25),
        ("Account_Type", 20),
        ("Last_Test_Date", 16),
        ("Test_Required_By", 16),
        ("Test_Status", 15),
        ("Current_Status", 15),
        ("Credential_Current", 15),
        ("Logging_Verified", 15),
        ("Readiness_Score", 15),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_status = DataValidation(type="list", formula1='"Disabled,Enabled,Unknown"', allow_blank=False)
    ws.add_data_validation(dv_status)

    dv_bool = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv_bool)

    for r in range(4, 24):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        # Test required by formula (1 year from last test)
        ws.cell(row=r, column=5).value = f'=IF(D{r}<>"",DATE(YEAR(D{r})+1,MONTH(D{r}),DAY(D{r})),"")'
        ws.cell(row=r, column=5).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Test status formula
        ws.cell(row=r, column=6).value = f'=IF(E{r}<>"",IF(E{r}<TODAY(),"Overdue","Current"),"")'
        ws.cell(row=r, column=6).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        dv_status.add(ws.cell(row=r, column=7))
        dv_bool.add(ws.cell(row=r, column=8))
        dv_bool.add(ws.cell(row=r, column=9))

        # Readiness score formula
        ws.cell(row=r, column=10).value = f'=IF(A{r}<>"",(IF(F{r}="Current",30,0)+IF(G{r}="Disabled",30,0)+IF(H{r}="Yes",20,0)+IF(I{r}="Yes",20,0)),"")'
        ws.cell(row=r, column=10).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.freeze_panes = "C4"


# =============================================================================
# PERSONNEL STATUS SHEET
# =============================================================================
def create_personnel_status_sheet(ws, styles):
    """Create Personnel_Status sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "PERSONNEL STATUS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Role_ID", 12),
        ("Role_Name", 30),
        ("Primary_Assigned", 15),
        ("Backup1_Assigned", 15),
        ("Backup2_Assigned", 15),
        ("Cross_Training", 15),
        ("Last_Contact_Test", 16),
        ("Contact_Test_Status", 18),
        ("Last_Drill", 16),
        ("Drill_Status", 15),
        ("Availability_Score", 15),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_bool = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv_bool)

    dv_training = DataValidation(type="list", formula1='"Complete,Partial,None"', allow_blank=False)
    ws.add_data_validation(dv_training)

    for r in range(4, 24):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        dv_bool.add(ws.cell(row=r, column=3))
        dv_bool.add(ws.cell(row=r, column=4))
        dv_bool.add(ws.cell(row=r, column=5))
        dv_training.add(ws.cell(row=r, column=6))

        # Contact test status (quarterly = 90 days)
        ws.cell(row=r, column=8).value = f'=IF(G{r}<>"",IF(G{r}+90<TODAY(),"Overdue","Current"),"")'
        ws.cell(row=r, column=8).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Drill status (semi-annual = 180 days)
        ws.cell(row=r, column=10).value = f'=IF(I{r}<>"",IF(I{r}+180<TODAY(),"Overdue","Current"),"")'
        ws.cell(row=r, column=10).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.freeze_panes = "C4"


# =============================================================================
# SECURITY DEBT STATUS SHEET
# =============================================================================
def create_security_debt_status_sheet(ws, styles):
    """Create Security_Debt_Status sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "SECURITY DEBT STATUS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Period", 15),
        ("Total_Open_Items", 18),
        ("Items_0_30_Days", 18),
        ("Items_31_60_Days", 18),
        ("Items_61_90_Days", 18),
        ("Items_Over_90_Days", 18),
        ("Items_Closed", 15),
        ("Closure_Rate", 15),
        ("Escalations", 12),
        ("Executive_Escalations", 18),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(4, 16):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        # Closure rate formula
        ws.cell(row=r, column=8).value = f'=IF(B{r}+G{r}>0,G{r}/(B{r}+G{r}),"")'
        ws.cell(row=r, column=8).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws.cell(row=r, column=8).number_format = '0%'

    ws.freeze_panes = "B4"


# =============================================================================
# DISRUPTION HISTORY SHEET
# =============================================================================
def create_disruption_history_sheet(ws, styles):
    """Create Disruption_History sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "DISRUPTION HISTORY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Incident_ID", 15),
        ("Incident_Date", 16),
        ("Incident_Type", 25),
        ("Duration_Hours", 15),
        ("Security_Posture_Level", 20),
        ("Emergency_Access_Used", 18),
        ("Security_Incidents_During", 20),
        ("Security_Compromises", 18),
        ("Recovery_Phase_Completed", 22),
        ("Lessons_Learned_Complete", 20),
        ("Security_Outcome", 18),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_posture = DataValidation(type="list", formula1='"Normal,Elevated,Degraded,Emergency"', allow_blank=False)
    ws.add_data_validation(dv_posture)

    dv_bool = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv_bool)

    dv_phase = DataValidation(
        type="list",
        formula1='"Immediate,Short-term,Medium-term,Long-term,Not Completed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_phase)

    dv_outcome = DataValidation(type="list", formula1='"No Issues,Minor Issues,Major Issues"', allow_blank=False)
    ws.add_data_validation(dv_outcome)

    for r in range(4, 24):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        dv_posture.add(ws.cell(row=r, column=5))
        dv_bool.add(ws.cell(row=r, column=6))
        dv_bool.add(ws.cell(row=r, column=8))
        dv_phase.add(ws.cell(row=r, column=9))
        dv_bool.add(ws.cell(row=r, column=10))
        dv_outcome.add(ws.cell(row=r, column=11))

    ws.freeze_panes = "C4"


# =============================================================================
# TREND ANALYSIS SHEET
# =============================================================================
def create_trend_analysis_sheet(ws, styles):
    """Create Trend_Analysis sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "TREND ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Period", 15),
        ("BCDR_Coverage_Pct", 18),
        ("Emergency_Test_Pct", 18),
        ("Personnel_Availability_Pct", 22),
        ("Security_Debt_Over90", 20),
        ("Disruptions_Count", 18),
        ("Security_Incidents_Count", 20),
        ("Overall_Score", 15),
        ("Trend", 12),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate with quarters
    periods = ["2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4", "2026-Q1"]

    for row_idx, period in enumerate(periods, start=4):
        ws.cell(row=row_idx, column=1, value=period)
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        # Format percentage columns
        for c in [2, 3, 4]:
            ws.cell(row=row_idx, column=c).number_format = '0%'

        # Overall score formula
        ws.cell(row=row_idx, column=8).value = f'=IF(B{row_idx}<>"",AVERAGE(B{row_idx}:D{row_idx}),"")'
        ws.cell(row=row_idx, column=8).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws.cell(row=row_idx, column=8).number_format = '0%'

    # Add more blank rows
    for r in range(len(periods) + 4, len(periods) + 12):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    ws.freeze_panes = "B4"


# =============================================================================
# DATA SOURCES SHEET
# =============================================================================
def create_data_sources_sheet(ws, styles):
    """Create Data_Sources sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "DATA SOURCES"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Source_ID", 12),
        ("Source_Name", 35),
        ("Source_Type", 18),
        ("File_Path", 50),
        ("Sheet_Name", 25),
        ("Last_Updated", 16),
        ("Update_Frequency", 18),
        ("Data_Owner", 25),
        ("Link_Status", 15),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate with expected sources
    sources = [
        ("DS-001", "Security Controls During Disruption", "Workbook", "ISMS-IMP-A.5.29.1_*.xlsx", "BCDR_Security_Review"),
        ("DS-002", "Degraded Mode Requirements", "Workbook", "ISMS-IMP-A.5.29.2_*.xlsx", "BreakGlass_Accounts"),
        ("DS-003", "Degraded Mode Requirements", "Workbook", "ISMS-IMP-A.5.29.2_*.xlsx", "Personnel_Availability"),
        ("DS-004", "Degraded Mode Requirements", "Workbook", "ISMS-IMP-A.5.29.2_*.xlsx", "Security_Debt_Register"),
        ("DS-005", "Recovery Verification", "Workbook", "ISMS-IMP-A.5.29.3_*.xlsx", "Lessons_Learned"),
    ]

    dv_type = DataValidation(type="list", formula1='"Workbook,Database,Manual Entry"', allow_blank=False)
    ws.add_data_validation(dv_type)

    dv_status = DataValidation(type="list", formula1='"Active,Broken,Not Linked"', allow_blank=False)
    ws.add_data_validation(dv_status)

    for row_idx, source in enumerate(sources, start=4):
        for col_idx, value in enumerate(source, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]

        for c in range(6, 10):
            ws.cell(row=row_idx, column=c).fill = styles["input_cell"]["fill"]

        dv_type.add(ws.cell(row=row_idx, column=3))
        dv_status.add(ws.cell(row=row_idx, column=9))

    for r in range(len(sources) + 4, len(sources) + 14):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_type.add(ws.cell(row=r, column=3))
        dv_status.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "C4"


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Dashboard Type", "Compliance Dashboard"),
        ("Control Reference", CONTROL_ID),
        ("Version", "1.0"),
        ("Generated Date", GENERATED_DATE),
        ("Update Frequency", "Monthly"),
        ("Owner", "CISO"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1

    row += 1
    ws[f"A{row}"] = "HOW TO USE THIS DASHBOARD"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Update Data_Sources with links to source workbooks",
        "2. Refresh BCDR_Security_Status from ISMS-IMP-A.5.29.1",
        "3. Refresh Emergency_Access_Status from ISMS-IMP-A.5.29.2",
        "4. Refresh Personnel_Status from ISMS-IMP-A.5.29.2",
        "5. Update Security_Debt_Status monthly",
        "6. Record Disruption_History after each incident",
        "7. Update Trend_Analysis quarterly",
        "8. Review and update Executive_Dashboard",
        "9. Obtain monthly CISO approval",
    ]

    row += 1
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50
    ws.freeze_panes = "A4"


# =============================================================================
# APPROVAL SIGN-OFF SHEET
# =============================================================================
def create_approval_signoff_sheet(ws, styles):
    """Create Approval_SignOff sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "DASHBOARD APPROVAL"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Dashboard Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Dashboard Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Reporting Period", ""),
        ("Generated Date", GENERATED_DATE),
        ("Overall Compliance Score", ""),
        ("Key Issues Count", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "PREPARED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    for field in ["Name", "Role", "Date"]:
        row += 1
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (Security Manager)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    for field in ["Name", "Date", "Signature"]:
        row += 1
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    for field in ["Name", "Date", "Approval Status", "Comments"]:
        row += 1
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A3"


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/10] Creating Executive_Dashboard sheet...")
        create_executive_dashboard_sheet(wb["Executive_Dashboard"], styles)

        logger.info("[2/10] Creating BCDR_Security_Status sheet...")
        create_bcdr_security_status_sheet(wb["BCDR_Security_Status"], styles)

        logger.info("[3/10] Creating Emergency_Access_Status sheet...")
        create_emergency_access_status_sheet(wb["Emergency_Access_Status"], styles)

        logger.info("[4/10] Creating Personnel_Status sheet...")
        create_personnel_status_sheet(wb["Personnel_Status"], styles)

        logger.info("[5/10] Creating Security_Debt_Status sheet...")
        create_security_debt_status_sheet(wb["Security_Debt_Status"], styles)

        logger.info("[6/10] Creating Disruption_History sheet...")
        create_disruption_history_sheet(wb["Disruption_History"], styles)

        logger.info("[7/10] Creating Trend_Analysis sheet...")
        create_trend_analysis_sheet(wb["Trend_Analysis"], styles)

        logger.info("[8/10] Creating Data_Sources sheet...")
        create_data_sources_sheet(wb["Data_Sources"], styles)

        logger.info("[9/10] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[10/10] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.29.4 specification
# =============================================================================
