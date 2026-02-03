#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.29.3 - Recovery Security Verification Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.29: Information Security During Disruption
Assessment Domain 3 of 4: Recovery Security Verification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates an Excel workbook for documenting security verification
activities during and after recovery from disruption.

**Generated Workbook Structure:**
1. Instructions - Assessment guidance
2. Recovery_Checklist - Phase-based validation activities
3. Emergency_Access_Closure - Deactivation verification
4. Control_Validation - Security control restoration
5. Anomaly_Detection - Post-disruption log analysis
6. Security_Debt_Closure - Debt item resolution tracking
7. Lessons_Learned - Security review findings
8. Evidence_Register - Audit evidence tracking
9. Approval_SignOff - Phase and final sign-off

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.29.3"
WORKBOOK_NAME = "Recovery Security Verification"
CONTROL_ID = "A.5.29"
CONTROL_NAME = "Information Security During Disruption"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "complete": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
    }


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Recovery_Checklist",
        "Emergency_Access_Closure",
        "Control_Validation",
        "Anomaly_Detection",
        "Security_Debt_Closure",
        "Lessons_Learned",
        "Evidence_Register",
        "Approval_SignOff",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Recovery Security Verification"),
        ("Control Reference", CONTROL_ID),
        ("Version", "1.0"),
        ("Disruption Incident ID", ""),
        ("Disruption Start Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "RECOVERY PHASES"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    phases = [
        ("Immediate (0-24h)", "Critical stabilisation, disable emergency access, initial log review"),
        ("Short-term (1-7d)", "Full control validation, vulnerability scan, access review"),
        ("Medium-term (1-4w)", "Security debt closure, deferred patches, access recertification"),
        ("Long-term (1-3m)", "Lessons learned, policy updates, training updates"),
    ]

    row += 1
    for phase, desc in phases:
        ws[f"A{row}"] = phase
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = desc
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    ws.freeze_panes = "A4"


# =============================================================================
# RECOVERY CHECKLIST SHEET
# =============================================================================
def create_recovery_checklist_sheet(ws, styles):
    """Create Recovery_Checklist sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "RECOVERY CHECKLIST - Phase-Based Security Validation"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Checklist_ID", 15),
        ("Phase", 15),
        ("Activity", 45),
        ("Responsible_Party", 25),
        ("Target_Completion", 20),
        ("Status", 15),
        ("Completion_Date", 16),
        ("Completed_By", 25),
        ("Evidence_Reference", 20),
        ("Notes", 40),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate checklist items
    checklist = [
        ("RC-001", "Immediate", "Verify critical security controls operational", "Security Team", "Within 4 hours"),
        ("RC-002", "Immediate", "Disable all emergency access accounts", "Security Team", "Within 8 hours"),
        ("RC-003", "Immediate", "Rotate credentials for activated break-glass accounts", "IAM Team", "Within 12 hours"),
        ("RC-004", "Immediate", "Initial log review for anomalies during disruption", "SOC", "Within 24 hours"),
        ("RC-005", "Immediate", "Notify CISO of emergency access deactivation", "Security Team", "Within 24 hours"),
        ("RC-006", "Short-term", "Complete security control validation checklist", "Security Team", "Within 3 days"),
        ("RC-007", "Short-term", "Execute vulnerability scan on recovered systems", "Security Team", "Within 5 days"),
        ("RC-008", "Short-term", "Review access changes during disruption period", "Security Team", "Within 5 days"),
        ("RC-009", "Short-term", "Complete detailed log analysis for disruption period", "SOC", "Within 7 days"),
        ("RC-010", "Short-term", "Security Manager sign-off on short-term phase", "Security Manager", "Within 7 days"),
        ("RC-011", "Medium-term", "Close all security debt items from disruption", "Security Team", "Within 2 weeks"),
        ("RC-012", "Medium-term", "Apply all deferred patches", "IT Operations", "Within 3 weeks"),
        ("RC-013", "Medium-term", "Complete full access recertification", "Security Team", "Within 4 weeks"),
        ("RC-014", "Medium-term", "CISO sign-off on medium-term phase", "CISO", "Within 4 weeks"),
        ("RC-015", "Long-term", "Conduct lessons learned security review", "CISO", "Within 6 weeks"),
        ("RC-016", "Long-term", "Update BC/DR plans with security lessons", "Security Team", "Within 8 weeks"),
        ("RC-017", "Long-term", "Update security awareness training", "Security Team", "Within 10 weeks"),
        ("RC-018", "Long-term", "Final CISO sign-off for return to normal", "CISO", "Within 12 weeks"),
    ]

    dv_phase = DataValidation(
        type="list",
        formula1='"Immediate,Short-term,Medium-term,Long-term"',
        allow_blank=False
    )
    ws.add_data_validation(dv_phase)

    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Complete,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for row_idx, item in enumerate(checklist, start=4):
        for col_idx, value in enumerate(item, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            if col_idx == 3:
                cell.alignment = Alignment(wrap_text=True, vertical="center")

        dv_phase.add(ws.cell(row=row_idx, column=2))
        for c in range(6, 11):
            ws.cell(row=row_idx, column=c).fill = styles["input_cell"]["fill"]
        dv_status.add(ws.cell(row=row_idx, column=6))

    for r in range(len(checklist) + 4, len(checklist) + 14):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_phase.add(ws.cell(row=r, column=2))
        dv_status.add(ws.cell(row=r, column=6))

    ws.freeze_panes = "D4"


# =============================================================================
# EMERGENCY ACCESS CLOSURE SHEET
# =============================================================================
def create_emergency_access_closure_sheet(ws, styles):
    """Create Emergency_Access_Closure sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = "EMERGENCY ACCESS CLOSURE - Deactivation Verification"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Account_ID", 15),
        ("Account_Name", 25),
        ("Activation_Date", 16),
        ("Deactivation_Date", 16),
        ("Deactivated_By", 25),
        ("Credential_Rotated", 15),
        ("Rotation_Date", 16),
        ("Usage_Reviewed", 15),
        ("Review_Date", 16),
        ("Anomalies_Found", 15),
        ("Anomaly_Details", 40),
        ("CISO_Notified", 15),
        ("Verification_Status", 15),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_bool = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv_bool)

    dv_status = DataValidation(type="list", formula1='"Pending,Verified"', allow_blank=False)
    ws.add_data_validation(dv_status)

    for r in range(4, 24):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        dv_bool.add(ws.cell(row=r, column=6))
        dv_bool.add(ws.cell(row=r, column=8))
        dv_bool.add(ws.cell(row=r, column=10))
        dv_bool.add(ws.cell(row=r, column=12))
        dv_status.add(ws.cell(row=r, column=13))

    ws.freeze_panes = "C4"


# =============================================================================
# CONTROL VALIDATION SHEET
# =============================================================================
def create_control_validation_sheet(ws, styles):
    """Create Control_Validation sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "CONTROL VALIDATION - Security Control Restoration"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Control_ID", 15),
        ("Control_Name", 30),
        ("Category", 20),
        ("Pre_Disruption_Status", 18),
        ("Current_Status", 18),
        ("Validation_Method", 30),
        ("Validation_Date", 16),
        ("Validated_By", 25),
        ("Gaps_Identified", 40),
        ("Remediation_Required", 15),
        ("Remediation_Date", 16),
        ("Remediation_Status", 15),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_status = DataValidation(
        type="list",
        formula1='"Operational,Partial,Not Operational"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_bool = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv_bool)

    dv_remediation = DataValidation(type="list", formula1='"Open,Closed"', allow_blank=False)
    ws.add_data_validation(dv_remediation)

    for r in range(4, 54):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        dv_status.add(ws.cell(row=r, column=5))
        dv_bool.add(ws.cell(row=r, column=10))
        dv_remediation.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "C4"


# =============================================================================
# ANOMALY DETECTION SHEET
# =============================================================================
def create_anomaly_detection_sheet(ws, styles):
    """Create Anomaly_Detection sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = "ANOMALY DETECTION - Post-Disruption Log Analysis"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Anomaly_ID", 12),
        ("Detection_Source", 25),
        ("Detection_Date", 16),
        ("Anomaly_Type", 20),
        ("Description", 50),
        ("Severity", 12),
        ("Related_Timeframe", 25),
        ("Investigation_Status", 20),
        ("Investigator", 25),
        ("Findings", 50),
        ("Actions_Taken", 40),
        ("Escalated", 10),
        ("Escalated_To", 25),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_source = DataValidation(
        type="list",
        formula1='"SIEM,Log Review,User Report,Automated Alert,Manual Analysis"',
        allow_blank=False
    )
    ws.add_data_validation(dv_source)

    dv_type = DataValidation(
        type="list",
        formula1='"Authentication,Access,Network,Data,Configuration,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_severity = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low,Informational"',
        allow_blank=False
    )
    ws.add_data_validation(dv_severity)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,Investigating,Closed - False Positive,Closed - Confirmed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_bool = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv_bool)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"AN-{r-3:03d}").font = Font(color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        dv_source.add(ws.cell(row=r, column=2))
        dv_type.add(ws.cell(row=r, column=4))
        dv_severity.add(ws.cell(row=r, column=6))
        dv_status.add(ws.cell(row=r, column=8))
        dv_bool.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "E4"


# =============================================================================
# SECURITY DEBT CLOSURE SHEET
# =============================================================================
def create_security_debt_closure_sheet(ws, styles):
    """Create Security_Debt_Closure sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = "SECURITY DEBT CLOSURE - Remediation Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Debt_ID", 15),
        ("Debt_Type", 20),
        ("Description", 45),
        ("Created_Date", 16),
        ("Original_Target", 16),
        ("Remediation_Action", 45),
        ("Action_Date", 16),
        ("Action_By", 25),
        ("Verification_Method", 35),
        ("Verification_Date", 16),
        ("Verified_By", 25),
        ("Closure_Status", 15),
        ("Evidence_Reference", 20),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_status = DataValidation(
        type="list",
        formula1='"Pending,Verified,Closed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 34):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        dv_status.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "D4"


# =============================================================================
# LESSONS LEARNED SHEET
# =============================================================================
def create_lessons_learned_sheet(ws, styles):
    """Create Lessons_Learned sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "LESSONS LEARNED - Security Review Findings"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Finding_ID", 12),
        ("Category", 20),
        ("Finding_Description", 50),
        ("Impact", 40),
        ("Root_Cause", 40),
        ("Recommendation", 50),
        ("Action_Type", 20),
        ("Action_Owner", 25),
        ("Target_Date", 16),
        ("Status", 12),
        ("Completion_Date", 16),
        ("Evidence_Reference", 20),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_category = DataValidation(
        type="list",
        formula1='"Planning,Controls,Personnel,Communication,Tools,Process,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_category)

    dv_action = DataValidation(
        type="list",
        formula1='"Policy Update,Procedure Update,Training,Tool Enhancement,Process Change"',
        allow_blank=False
    )
    ws.add_data_validation(dv_action)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Closed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 24):
        ws.cell(row=r, column=1, value=f"LL-{r-3:03d}").font = Font(color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        dv_category.add(ws.cell(row=r, column=2))
        dv_action.add(ws.cell(row=r, column=7))
        dv_status.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "D4"


# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================
def create_evidence_register_sheet(ws, styles):
    """Create Evidence_Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 45),
        ("Related_Section", 25),
        ("Collection_Date", 16),
        ("Location", 40),
        ("Collected_By", 25),
        ("Valid_Until", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Document,Approval,Test Result,Configuration,Screenshot,Attestation"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "C4"


# =============================================================================
# APPROVAL SIGN-OFF SHEET
# =============================================================================
def create_approval_signoff_sheet(ws, styles):
    """Create Approval_SignOff sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "RECOVERY PHASE SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Recovery Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Disruption Incident", ""),
        ("Recovery Checklist Items Complete", '=COUNTIF(Recovery_Checklist!F4:F33,"Complete")'),
        ("Emergency Access Accounts Verified", '=COUNTIF(Emergency_Access_Closure!M4:M23,"Verified")'),
        ("Controls Validated", '=COUNTIF(Control_Validation!E4:E53,"Operational")'),
        ("Open Anomalies", '=COUNTIF(Anomaly_Detection!H4:H53,"Open")+COUNTIF(Anomaly_Detection!H4:H53,"Investigating")'),
        ("Security Debt Closed", '=COUNTIF(Security_Debt_Closure!L4:L33,"Closed")'),
        ("Lessons Learned Actions Open", '=COUNTIF(Lessons_Learned!J4:J23,"Open")+COUNTIF(Lessons_Learned!J4:J23,"In Progress")'),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "" or not str(value).startswith("="):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    # Phase sign-offs
    phases = [
        ("IMMEDIATE PHASE SIGN-OFF", "Security Team Lead"),
        ("SHORT-TERM PHASE SIGN-OFF", "Security Manager"),
        ("MEDIUM-TERM PHASE SIGN-OFF", "CISO"),
        ("LONG-TERM PHASE / RETURN TO NORMAL", "CISO"),
    ]

    for phase_name, approver in phases:
        row += 2
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = phase_name
        ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

        for field in [f"Approver ({approver})", "Date", "Status", "Comments"]:
            row += 1
            ws[f"A{row}"] = field + ":"
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40
    ws.freeze_panes = "A3"


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/9] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/9] Creating Recovery_Checklist sheet...")
        create_recovery_checklist_sheet(wb["Recovery_Checklist"], styles)

        logger.info("[3/9] Creating Emergency_Access_Closure sheet...")
        create_emergency_access_closure_sheet(wb["Emergency_Access_Closure"], styles)

        logger.info("[4/9] Creating Control_Validation sheet...")
        create_control_validation_sheet(wb["Control_Validation"], styles)

        logger.info("[5/9] Creating Anomaly_Detection sheet...")
        create_anomaly_detection_sheet(wb["Anomaly_Detection"], styles)

        logger.info("[6/9] Creating Security_Debt_Closure sheet...")
        create_security_debt_closure_sheet(wb["Security_Debt_Closure"], styles)

        logger.info("[7/9] Creating Lessons_Learned sheet...")
        create_lessons_learned_sheet(wb["Lessons_Learned"], styles)

        logger.info("[8/9] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[9/9] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.29.3 specification
# =============================================================================
