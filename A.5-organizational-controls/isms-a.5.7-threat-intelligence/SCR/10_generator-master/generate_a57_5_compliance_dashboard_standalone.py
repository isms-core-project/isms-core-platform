#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.7.5 - Threat Intelligence Standalone Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.7: Threat Intelligence
Assessment Domain 5 of 5: Standalone Compliance Dashboard (Executive Reporting)

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific executive reporting requirements, compliance
thresholds, and stakeholder communication needs.

Key customization areas:
1. Executive KPIs and metrics (align with management reporting needs)
2. Compliance scoring thresholds (adapt to audit requirements)
3. Dashboard layout and visualizations (match executive preferences)
4. Manual data entry fields (customize for your operational context)
5. Regulatory reporting requirements (jurisdiction-specific)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.7 Threat Intelligence Framework

⚠️ IMPORTANT DISTINCTION: This is a STANDALONE dashboard that does NOT use
external workbook references. It relies on manual data entry for executive
reporting when assessment workbooks are not available or appropriate for
distribution (e.g., board meetings, external audits, regulatory reporting).

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a self-contained Excel dashboard for executive reporting
and compliance validation without dependencies on the detailed assessment
workbooks (A.5.7.1-3). It provides high-level compliance status, key metrics,
and audit evidence in a standalone format suitable for management briefings,
board presentations, and regulatory submissions.

**Purpose:**
Enables executive oversight and external reporting of threat intelligence
program effectiveness without exposing detailed operational assessments or
requiring access to the full assessment workbook suite (A.5.7.1-3).

**Use Cases:**
1. Board of Directors presentations (quarterly/annual security reports)
2. Executive management briefings (CISO reporting to CEO/COO)
3. External audit submissions (when full workbooks contain excessive detail)
4. Regulatory reporting (compliance status to regulators)
5. Insurance underwriting (cyber insurance policy applications)
6. M&A due diligence (security program maturity demonstration)
7. Client security questionnaires (SOC 2, ISO audit evidence)

**Key Distinction from A.5.7.4:**
- A.5.7.4 Dashboard: Uses external Excel references, consolidates real-time data
  from detailed assessment workbooks, designed for internal operational oversight
  
- A.5.7.5 Dashboard (THIS): Standalone with manual data entry, no external
  dependencies, designed for executive/external reporting with controlled detail

**Assessment Scope:**
- Overall Control A.5.7 compliance status (high-level summary)
- Source portfolio health (key metrics only, not full inventory)
- Collection/analysis effectiveness (workflow maturity, not detailed metrics)
- Integration/distribution performance (operational effectiveness indicators)
- VulnerabilityThreatLink (VTL) workflow status (executive summary)
- CVSS quality indicators (accuracy percentages, not source-level detail)
- MITRE ATT&CK coverage (summary percentages by tactic)
- Key performance indicators (KPIs) suitable for executive review
- Compliance status and audit readiness (traffic light indicators)
- Strategic gaps and remediation priorities (high-level only)

**Generated Dashboard Structure (8 Sheets):**
1. Executive_Summary - One-page compliance status for executives
2. Source_Portfolio_Summary - High-level source health metrics
3. Workflow_Effectiveness - Collection, analysis, distribution maturity
4. VTL_Integration_Status - VulnerabilityThreatLink health (simplified)
5. CVSS_Quality_Summary - CVSS accuracy and coverage (percentages)
6. MITRE_Coverage_Summary - ATT&CK coverage by tactic (high-level)
7. Compliance_Status - ISO 27001 A.5.7 requirement checklist
8. Metadata_Approvals - Assessment metadata and executive sign-off

**Key Features:**
- MANUAL data entry fields (no external workbook dependencies)
- Executive-appropriate summary metrics (no operational detail)
- Traffic light status indicators for quick visual assessment
- Compliance checklist aligned with ISO 27001:2022 A.5.7 requirements
- VTL workflow summary (demonstrates operational value)
- Clear instructions for completing manual data entry fields
- Protected formulas with unprotected input cells
- Single-file distribution (no broken references risk)
- Suitable for external stakeholder review
- Executive sign-off workflow

**VulnerabilityThreatLink (VTL) Executive Summary:**
Sheet 4 (VTL_Integration_Status) provides simplified VTL metrics suitable
for executive understanding without overwhelming technical detail:

Executive VTL Metrics:
- VTL Workflow Status: Operational / Degraded / Not Implemented
- CVSS Coverage: % of threat intelligence with severity ratings
- CVSS Accuracy: % accuracy vs. authoritative sources
- Emergency Patch Rate: % of exploited CVEs auto-escalated
- Average Response Time: Days from exploitation to patch deployment

This demonstrates operational security value in business terms.

**Integration:**
This standalone dashboard is INDEPENDENT from:
- A.5.7.1 Sources Assessment (data entered manually, not referenced)
- A.5.7.2 Collection & Analysis Assessment (data entered manually)
- A.5.7.3 Integration & Distribution Assessment (data entered manually)
- A.5.7.4 Effectiveness Dashboard (optional data source for manual entry)

Data can be manually transcribed from A.5.7.4 dashboard or entered directly
based on operational knowledge. No external references = no broken links.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

No Input Files Required:
    This script generates a standalone workbook with no external dependencies.
    All data is entered manually into the generated workbook.

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a57_5_compliance_dashboard_standalone.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a57_5_compliance_dashboard_standalone.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a57_5_compliance_dashboard_standalone.py --date 20250115

Output:
    File: ISMS_A_5_7_5_Standalone_Dashboard_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)
    Sheets: 8 worksheets (see structure above)

Post-Generation Steps:
    1. Read instructions on each sheet carefully
    2. Manually enter or transcribe data from A.5.7.4 dashboard (if available)
    3. Complete all yellow-highlighted input fields
    4. Review calculated compliance status and traffic light indicators
    5. Verify VTL workflow status accurately reflects operations
    6. Validate CVSS quality metrics match operational reality
    7. Review MITRE ATT&CK coverage percentages for accuracy
    8. Complete compliance checklist (Sheet 7) with evidence links
    9. Obtain executive sign-off (Sheet 8 - Metadata_Approvals)
    10. Distribute to intended audience (board, auditors, regulators)
    11. Archive signed version for audit trail

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.7
Assessment Domain:    5 of 5 (Standalone Dashboard - Executive Reporting)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.7: Threat Intelligence Policy (Governance)
    - ISMS-IMP-A.5.7.5: Standalone Compliance Dashboard Implementation Guide
    - ISMS-IMP-A.5.7.4: Threat Intelligence Effectiveness Dashboard (data source)
    - ISMS-IMP-A.5.7.1: Threat Intelligence Sources Assessment (reference)
    - ISMS-IMP-A.5.7.2: Intelligence Collection & Analysis Assessment (reference)
    - ISMS-IMP-A.5.7.3: Intelligence Integration & Distribution Assessment (reference)
    - ISO 27002:2022 Implementation Guidance for Control A.5.7
    - ISO/IEC 27001:2022 Annex A Control Requirements

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
- Initial standalone dashboard implementation (previous versions used external refs)
    - Designed for executive/external reporting without operational detail
    - Added VTL workflow summary sheet (simplified for executive understanding)
    - Enhanced compliance checklist aligned with ISO 27001:2022 requirements
    - Added manual data entry instructions on all input sheets
    - Improved traffic light status indicators for visual communication
    - Added executive sign-off workflow (Sheet 8)
    - Removed all external workbook dependencies

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Standalone vs. Consolidated Dashboard:**
Organizations implementing Control A.5.7 should maintain BOTH dashboards:

A.5.7.4 (Consolidated Dashboard):
- For: CISO, threat intel team, ISMS team, internal auditors
- Contains: Detailed metrics, operational data, full assessment consolidation
- Uses: External workbook references for real-time updates
- Purpose: Operational oversight and internal compliance management

A.5.7.5 (Standalone Dashboard - THIS):
- For: Board, executives, external auditors, regulators, clients
- Contains: Executive summary metrics, high-level compliance status
- Uses: Manual data entry, no external dependencies
- Purpose: External reporting and executive communication

Use the right tool for the right audience.

**Manual Data Entry Accuracy:**
Because this dashboard relies on manual data entry rather than automated
consolidation, accuracy depends on the person completing the fields.

Best Practices:
- Designate a single person to complete the dashboard (avoid multiple editors)
- Transcribe data from A.5.7.4 dashboard where available
- Cross-check entered metrics against operational reality
- Have a second reviewer validate before executive sign-off
- Document data sources and assumptions in comments
- Update quarterly to reflect current program status

Inaccurate self-assessment undermines credibility with auditors and executives.

**VTL Executive Communication:**
Executives and board members may not understand technical VTL details.
Frame VTL metrics in business terms:

Technical: "CVSS accuracy 92%, emergency patch SLA 24h"
Executive: "Threat intelligence enables automated vulnerability prioritization,
            reducing critical vulnerability exposure from days to hours"

Focus on business risk reduction, not technical implementation details.

**Compliance Checklist Completion:**
Sheet 7 (Compliance_Status) maps ISO 27001:2022 A.5.7 requirements to
implementation evidence. For each requirement:
- Status: Implemented / Partially Implemented / Not Implemented
- Evidence: Document reference or location (e.g., "A.5.7.1 Sheet 14")
- Comments: Brief explanation of implementation approach

This checklist is CRITICAL for external audits. Incomplete checklists will
result in audit findings or corrective action requests.

**Audit Considerations:**
When using this standalone dashboard for audit evidence:
- Ensure manual data entry is ACCURATE (auditors will verify)
- Complete all compliance checklist items (Sheet 7)
- Obtain executive sign-off BEFORE audit (Sheet 8)
- Be prepared to provide detailed workbooks (A.5.7.1-3) if auditor requests
- Archive signed dashboard as official audit evidence

Auditors may request detailed workbooks for validation - don't destroy them.

**Data Protection:**
Even though this dashboard contains less detail than operational workbooks,
it still reveals sensitive information:
- Threat intelligence program maturity level
- Compliance gaps and deficiencies
- VTL workflow effectiveness (or lack thereof)
- Strategic improvement priorities

Classification: CONFIDENTIAL - Management Use Only. Control distribution
carefully based on recipient need-to-know.

**Maintenance:**
Update standalone dashboard:
- Quarterly: After A.5.7.4 dashboard updates (transcribe new metrics)
- Semi-annually: Before board/executive security program reviews
- Annually: Before external audit or ISO certification renewal
- Ad-hoc: Before regulatory reporting deadlines, M&A due diligence, etc.

Archive prior versions to demonstrate improvement over time.

**Quality Assurance:**
Before distributing to executives or external stakeholders:
- Verify all yellow-highlighted input fields are completed
- Validate calculated compliance percentages are accurate
- Check traffic light status indicators match narrative explanations
- Review compliance checklist completeness (no "TBD" entries)
- Proofread all text for clarity and professionalism
- Have CISO or ISMS manager review before executive sign-off

Poor quality dashboards damage credibility and trust.

**Regulatory Alignment:**
Some regulations require specific threat intelligence metrics in reporting:
- Financial sector: Fraud intelligence program maturity
- Healthcare: Healthcare-specific threat intelligence coverage
- Critical infrastructure: CISA information sharing participation
- Data protection: Privacy-preserving threat intelligence metrics

Customize dashboard manual entry fields to capture regulatory-specific metrics.

**Executive Sign-Off Workflow:**
Sheet 8 (Metadata_Approvals) includes executive sign-off fields:
- Prepared By: ISMS team member or threat intel lead
- Reviewed By: CISO or security director
- Approved By: Executive leadership (CIO, COO, or CEO)
- Board Acknowledgment: For board presentations

Obtain signatures BEFORE distribution to external parties. Unsigned dashboards
lack authority and may be questioned by auditors or regulators.

**Business Impact Communication:**
Use this dashboard to communicate threat intelligence business value:
- Risk Reduction: "VTL workflow reduced critical vulnerability exposure by 65%"
- Cost Avoidance: "Early threat detection prevented 3 potential incidents ($XXM)"
- Compliance: "ISO 27001 A.5.7 compliance achieved without corrective actions"
- Efficiency: "Automated threat intelligence reduced manual analysis by 40%"

Quantify security improvements in business terms executives understand.

**Distribution Control:**
Track dashboard distribution to external parties:
- Who received the dashboard (name, organization, role)
- When distributed (date)
- Purpose (audit, regulatory reporting, M&A due diligence, etc.)
- Restrictions (NDA, confidentiality terms)

Maintain distribution log for accountability and data protection compliance.

**No Broken References:**
Unlike A.5.7.4 dashboard, this standalone version will NEVER have broken
external references because it doesn't use them. This is the key advantage
for external distribution - recipients can open the file without errors.

**Historical Comparison:**
If you maintain quarterly standalone dashboards, you can create executive
trend reporting:
- Compliance score improvement over time
- VTL workflow maturity progression
- CVSS quality improvement trajectory
- Gap reduction progress

Archive prior quarters to enable trend analysis for board reporting.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)




# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.7.5"
WORKBOOK_NAME = "Threat Intelligence Standalone Compliance Dashboard"
CONTROL_ID = "A.5.7"
CONTROL_NAME = "Threat Intelligence"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with 9 sheets matching specification."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create 9 sheets as per ISMS-IMP-A.5.7.5 specification
    sheets = [
        "Instructions",
        "Monthly_Data_Entry",
        "Executive_Dashboard",
        "Trend_History",
        "Critical_Actions",
        "Quarterly_Summary",
        "Risk_Summary",
        "ROI_Summary",
        "Metadata",
    ]
    
    for name in sheets:
        wb.create_sheet(title=name)
    
    return wb


def get_styles():
    """Define all cell styles used throughout workbook."""
    return {
        # Headers
        "header_main": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            ),
        },
        "header_sub": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="2E75B5", end_color="2E75B5", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            ),
        },
        "section_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="003366"),
            "fill": PatternFill(start_color="D8E4F8", end_color="D8E4F8", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
        },
        
        # Data cells
        "data_normal": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC")
            ),
        },
        "data_center": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC")
            ),
        },
        "data_number": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="right", vertical="center"),
            "border": Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC")
            ),
        },
        
        # Special formatting
        "kpi_box": {
            "font": Font(name="Calibri", size=24, bold=True, color="003366"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "fill": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
            "border": Border(
                left=Side(style="medium"), right=Side(style="medium"),
                top=Side(style="medium"), bottom=Side(style="medium")
            ),
        },
        "critical_alert": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "warning": {
            "font": Font(name="Calibri", size=10, bold=True, color="000000"),
            "fill": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "success": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        
        # Instructions
        "instruction_text": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
        },
        "instruction_title": {
            "font": Font(name="Calibri", size=14, bold=True, color="003366"),
            "alignment": Alignment(horizontal="left", vertical="center"),
        },
    }


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = style_dict["font"]
    if "fill" in style_dict:
        cell.fill = style_dict["fill"]
    if "alignment" in style_dict:
        cell.alignment = style_dict["alignment"]
    if "border" in style_dict:
        cell.border = style_dict["border"]
    if "number_format" in style_dict:
        cell.number_format = style_dict["number_format"]


def set_column_widths(ws, widths):
    """Set column widths for a worksheet."""
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ============================================================================
# SECTION 2: SHEET 1 - INSTRUCTIONS
# ============================================================================

def create_instructions_sheet(wb):
    """Create Sheet 1: Instructions."""
    ws = wb["Instructions"]
    styles = get_styles()
    
    # Title
    ws["A1"] = "ISMS-IMP-A.5.7.5 - Threat Intelligence Standalone Dashboard"
    apply_style(ws["A1"], styles["instruction_title"])
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 20
    
    ws["A2"] = "Instructions & Monthly Workflow Guide"
    apply_style(ws["A2"], styles["header_main"])
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 25
    
    # Overview Section - Table Format
    row = 4
    ws[f"A{row}"] = "1. OVERVIEW"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 2
    ws[f"A{row}"] = "Purpose:"
    ws[f"B{row}"] = "Executive TI program dashboard for Board presentations and C-suite reporting"
    ws.merge_cells(f"B{row}:F{row}")
    
    row += 1
    ws[f"A{row}"] = "Audience:"
    ws[f"B{row}"] = "C-suite, Board members, External stakeholders"
    ws.merge_cells(f"B{row}:F{row}")
    
    row += 1
    ws[f"A{row}"] = "Update Time:"
    ws[f"B{row}"] = "10 minutes monthly + 30 minutes quarterly"
    ws.merge_cells(f"B{row}:F{row}")
    
    row += 1
    ws[f"A{row}"] = "Dependencies:"
    ws[f"B{row}"] = "NONE - Completely standalone (no external workbook links)"
    ws.merge_cells(f"B{row}:F{row}")
    apply_style(ws[f"B{row}"], styles["success"])
    
    # Key Difference Table
    row += 2
    ws[f"A{row}"] = "Key Difference: This Dashboard (5.7.5) vs Comprehensive Dashboard (5.7.4)"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 1
    headers = ["Feature", "5.7.4 Comprehensive", "5.7.5 Standalone (This)"]
    ws[f"A{row}"] = headers[0]
    ws[f"B{row}"] = headers[1]
    ws[f"D{row}"] = headers[2]
    apply_style(ws[f"A{row}"], styles["header_sub"])
    apply_style(ws[f"B{row}"], styles["header_sub"])
    apply_style(ws[f"D{row}"], styles["header_sub"])
    ws.merge_cells(f"B{row}:C{row}")
    ws.merge_cells(f"D{row}:F{row}")
    
    row += 1
    comparison_data = [
        ("Data Source", "External references (auto)", "Manual entry"),
        ("Update Time", "Seconds (refresh links)", "10 minutes"),
        ("Audience", "Security team, Auditors", "Board, C-suite"),
        ("Detail Level", "Operational", "Strategic"),
        ("Use Case", "Internal operations", "Executive briefings"),
    ]
    
    for feature, comp, standalone in comparison_data:
        ws[f"A{row}"] = feature
        ws[f"B{row}"] = comp
        ws[f"D{row}"] = standalone
        ws.merge_cells(f"B{row}:C{row}")
        ws.merge_cells(f"D{row}:F{row}")
        row += 1
    
    # Monthly Workflow - Clear Steps
    row += 1
    ws[f"A{row}"] = "2. MONTHLY WORKFLOW (10 Minutes) - Do on 1st Business Day"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 1
    ws[f"A{row}"] = "Step"
    ws[f"B{row}"] = "Action"
    ws[f"D{row}"] = "Details"
    apply_style(ws[f"A{row}"], styles["header_sub"])
    apply_style(ws[f"B{row}"], styles["header_sub"])
    apply_style(ws[f"D{row}"], styles["header_sub"])
    ws.merge_cells(f"D{row}:F{row}")
    
    row += 1
    workflow_steps = [
        ("1", "Open Sheet 2", "Monthly_Data_Entry sheet"),
        ("2", "Enter Period", "Format: January 2025"),
        ("3", "Fill Metrics", "20 fields across 6 sections (see Section 3 below)"),
        ("4", "Set Status", "Change dropdown to 'Final'"),
        ("5", "Check Dashboard", "Review Sheet 3 (auto-updates)"),
        ("6", "Update Actions", "Sheet 5 if needed"),
        ("7", "Update Risks", "Sheet 7 if needed"),
        ("8", "Save File", "Use dated filename for archive"),
        ("9", "Export PDF", "Sheet 3 → File → Save As → PDF"),
    ]
    
    for step, action, details in workflow_steps:
        ws[f"A{row}"] = step
        ws[f"B{row}"] = action
        ws[f"D{row}"] = details
        apply_style(ws[f"A{row}"], styles["data_center"])
        ws.merge_cells(f"B{row}:C{row}")
        ws.merge_cells(f"D{row}:F{row}")
        row += 1
    
    # Metric Sources - Simplified Table
    row += 1
    ws[f"A{row}"] = "3. WHERE TO FIND DATA (20 Metrics)"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"C{row}"] = "Data Source"
    apply_style(ws[f"A{row}"], styles["header_sub"])
    apply_style(ws[f"C{row}"], styles["header_sub"])
    ws.merge_cells(f"A{row}:B{row}")
    ws.merge_cells(f"C{row}:F{row}")
    
    row += 1
    metric_sources = [
        ("Active TI Sources", "ISMS-IMP-A.5.7.1 Sheet 2 or TI platform"),
        ("Average Source Quality", "Quality scores from evaluations (1-5 scale)"),
        ("Products Published", "Document repository count this month"),
        ("Stakeholder Satisfaction", "Monthly survey average (1-5 scale)"),
        ("IOCs Deployed", "SIEM dashboard count"),
        ("IOC Hit Rate %", "SIEM analytics: (Hits / Total) × 100"),
        ("CVSS 4.0 Adoption %", "Vuln mgmt: (v4.0 / Total) × 100"),
        ("High CVSS Open", "Count CVSS≥7.0 + active exploit (MUST=0)"),
        ("Incidents Prevented", "ISMS-IMP-A.5.7.3 Sheet 7"),
        ("Cost Avoidance CHF", "Sum of prevented incident costs"),
    ]
    
    for metric, source in metric_sources:
        ws[f"A{row}"] = metric
        ws[f"C{row}"] = source
        ws.merge_cells(f"A{row}:B{row}")
        ws.merge_cells(f"C{row}:F{row}")
        row += 1
    
    row += 1
    ws[f"A{row}"] = "See Sheet 2 for all 20 metrics with examples"
    ws.merge_cells(f"A{row}:F{row}")
    apply_style(ws[f"A{row}"], styles["instruction_text"])
    
    # Quarterly Tasks - Simple List
    row += 2
    ws[f"A{row}"] = "4. QUARTERLY TASKS (Additional 30 Min) - Do Last Week of Quarter"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 1
    ws[f"A{row}"] = "Task"
    ws[f"B{row}"] = "Action"
    apply_style(ws[f"A{row}"], styles["header_sub"])
    apply_style(ws[f"B{row}"], styles["header_sub"])
    ws.merge_cells(f"B{row}:F{row}")
    
    row += 1
    quarterly_tasks = [
        ("1", "Complete monthly workflow above"),
        ("2", "Sheet 6: Add quarterly highlights (top 3 achievements/challenges)"),
        ("3", "Sheet 8: Update ROI calculations"),
        ("4", "Export 3 PDFs: Sheets 3, 6, 8"),
        ("5", "Present to Board or distribute"),
    ]
    
    for task_num, action in quarterly_tasks:
        ws[f"A{row}"] = task_num
        ws[f"B{row}"] = action
        apply_style(ws[f"A{row}"], styles["data_center"])
        ws.merge_cells(f"B{row}:F{row}")
        row += 1
    
    # Contact Info - Simple
    row += 2
    ws[f"A{row}"] = "5. CONTACTS & RESOURCES"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 1
    contact_data = [
        ("CISO:", "[Name] - [Email]"),
        ("TI Team Lead:", "[Name] - [Email]"),
        ("ISMS Support:", "isms@[organization].com"),
        ("", ""),
        ("Policy:", "ISMS-POL-A.5.7"),
        ("Specification:", "ISMS-IMP-A.5.7.5"),
        ("Operational Dashboard:", "ISMS-IMP-A.5.7.4 (for internal use)"),
    ]
    
    for label, value in contact_data:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws.merge_cells(f"B{row}:F{row}")
        row += 1
    
    # Column widths
    set_column_widths(ws, [12, 20, 20, 20, 20, 20])
    
    logger.info("✅ Sheet 1: Instructions - Complete")


# ============================================================================
# SECTION 3: SHEET 2 - MONTHLY DATA ENTRY
# ============================================================================

def create_monthly_data_entry_sheet(wb):
    """Create Sheet 2: Monthly_Data_Entry - Primary data entry interface."""
    ws = wb["Monthly_Data_Entry"]
    styles = get_styles()
    
    # Title
    ws["A1"] = "Monthly Data Entry - Threat Intelligence Program"
    apply_style(ws["A1"], styles["header_main"])
    ws.merge_cells("A1:D1")
    
    ws["A2"] = "Complete all fields on 1st business day of new month (10 minutes)"
    apply_style(ws["A2"], styles["header_sub"])
    ws.merge_cells("A2:D2")
    
    # Section A: Document Control
    row = 4
    ws[f"A{row}"] = "A. DOCUMENT CONTROL"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:D{row}")
    
    row += 1
    headers = ["Field", "Value", "Instructions", "Example"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    # Document control fields
    row += 1
    doc_fields = [
        ("Reporting_Period", "", "Format: Month YYYY", "January 2025"),
        ("Entry_Date", "=TODAY()", "Auto-filled", "2025-02-03"),
        ("Entered_By", "", "Your email", "jane.ciso@example.com"),
        ("Status", "Draft", "Set to 'Final' when complete", "Draft / Final"),
    ]
    
    for field_name, formula, instruction, example in doc_fields:
        ws.cell(row=row, column=1, value=field_name)
        if formula:
            ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=3, value=instruction)
        ws.cell(row=row, column=4, value=example)
        row += 1
    
    # Status dropdown
    status_dv = DataValidation(type="list", formula1='"Draft,Final"', allow_blank=False)
    ws.add_data_validation(status_dv)
    status_dv.add(f"B{row-1}")
    
    # Section B: Source Portfolio Metrics
    row += 1
    ws[f"A{row}"] = "B. SOURCE PORTFOLIO METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:D{row}")
    
    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    row += 1
    source_fields = [
        ("Active_TI_Sources", "", "Count of active sources", "18"),
        ("Average_Source_Quality", "", "1.0-5.0 scale", "4.2"),
        ("Coverage_Gaps_Count", "", "Number of identified gaps", "2"),
        ("Annual_Source_Cost_CHF", "", "Total subscription costs", "450000"),
    ]
    
    for field_name, formula, instruction, example in source_fields:
        ws.cell(row=row, column=1, value=field_name)
        ws.cell(row=row, column=3, value=instruction)
        ws.cell(row=row, column=4, value=example)
        row += 1
    
    # Section C: Intelligence Production Metrics
    row += 1
    ws[f"A{row}"] = "C. INTELLIGENCE PRODUCTION METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:D{row}")
    
    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    row += 1
    production_fields = [
        ("Intelligence_Products_Published", "", "Count this month", "23"),
        ("Average_Time_To_Produce_Hours", "", "Hours from collection to publication", "6.5"),
        ("Stakeholder_Satisfaction_Rating", "", "1.0-5.0 scale", "4.3"),
        ("Intelligence_Consumption_Rate", "", "Percentage (0-100)", "82"),
    ]
    
    for field_name, formula, instruction, example in production_fields:
        ws.cell(row=row, column=1, value=field_name)
        ws.cell(row=row, column=3, value=instruction)
        ws.cell(row=row, column=4, value=example)
        row += 1
    
    # Section D: Integration & Effectiveness
    row += 1
    ws[f"A{row}"] = "D. INTEGRATION & EFFECTIVENESS METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:D{row}")
    
    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    row += 1
    integration_fields = [
        ("Security_Tools_Integrated", "", "Count of tools with TI feeds", "9"),
        ("IOCs_Deployed", "", "IOCs deployed this month", "1247"),
        ("IOC_Hit_Rate", "", "Percentage (0-100)", "12.3"),
        ("IOC_False_Positive_Rate", "", "Percentage (0-100)", "3.8"),
    ]
    
    for field_name, formula, instruction, example in integration_fields:
        ws.cell(row=row, column=1, value=field_name)
        ws.cell(row=row, column=3, value=instruction)
        ws.cell(row=row, column=4, value=example)
        row += 1
    
    # Section E: CVSS & Vulnerability Metrics
    row += 1
    ws[f"A{row}"] = "E. CVSS & VULNERABILITY METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:D{row}")
    
    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    row += 1
    cvss_fields = [
        ("CVSS_4_0_Adoption_Rate", "", "Percentage (0-100)", "68"),
        ("CVSS_Source_Accuracy_Rate", "", "Percentage (0-100)", "91"),
        ("High_CVSS_Active_Exploitation_Open", "", "Count (MUST be 0)", "0"),
        ("VTL_Records_Created", "", "Vulnerability-Threat Links", "47"),
    ]
    
    for field_name, formula, instruction, example in cvss_fields:
        ws.cell(row=row, column=1, value=field_name)
        ws.cell(row=row, column=3, value=instruction)
        ws.cell(row=row, column=4, value=example)
        row += 1
    
    # Section F: Prevention & Impact
    row += 1
    ws[f"A{row}"] = "F. PREVENTION & IMPACT METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:D{row}")
    
    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    row += 1
    prevention_fields = [
        ("Incidents_Prevented", "", "Count this month", "4"),
        ("Cost_Avoidance_CHF", "", "Estimated savings", "275000"),
        ("Risk_Assessments_Updated", "", "Count this month", "3"),
        ("P1_P2_Incidents_With_TI_Context", "", "Percentage (0-100)", "85"),
    ]
    
    for field_name, formula, instruction, example in prevention_fields:
        ws.cell(row=row, column=1, value=field_name)
        ws.cell(row=row, column=3, value=instruction)
        ws.cell(row=row, column=4, value=example)
        row += 1
    
    # Add notes section
    row += 2
    ws[f"A{row}"] = "NOTES / COMMENTS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    ws.merge_cells(f"A{row}:D{row+3}")
    
    # Column widths
    set_column_widths(ws, [35, 15, 40, 20])
    
    logger.info("✅ Sheet 2: Monthly_Data_Entry - Complete")


# This is Section 1 (Setup, Instructions, Data Entry)
# Continue in next message with remaining sheets...
# ============================================================================
# SECTION 4: SHEET 3 - EXECUTIVE DASHBOARD (ONE-PAGE VISUAL)
# ============================================================================

def create_executive_dashboard_sheet(wb):
    """Create Sheet 3: Executive_Dashboard - Auto-generated one-page summary."""
    ws = wb["Executive_Dashboard"]
    styles = get_styles()
    
    # Set print layout
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    
    # Header banner
    ws["A1"] = "[Organization Logo]"
    ws["E1"] = "THREAT INTELLIGENCE PROGRAM"
    apply_style(ws["E1"], styles["header_main"])
    ws.merge_cells("E1:J1")
    
    ws["A2"] = "Executive Summary Dashboard"
    apply_style(ws["A2"], styles["header_sub"])
    ws.merge_cells("A2:J2")
    
    # Reporting period and date
    ws["A3"] = "Reporting Period:"
    ws["B3"] = "=Monthly_Data_Entry!B5"  # Reference to reporting period
    ws["H3"] = "Report Date:"
    ws["I3"] = "=TODAY()"
    
    # Program Health Score (large composite metric)
    ws["A4"] = "Program Health Score"
    ws["A5"] = "=AVERAGE(B10,B11,B12,B13)"  # Will adjust based on actual formula
    apply_style(ws["A5"], styles["kpi_box"])
    ws.merge_cells("A5:B6")
    ws["A5"].number_format = "0.0"
    
    # KPI Boxes Row 1
    row = 8
    ws[f"A{row}"] = "KEY PERFORMANCE INDICATORS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:J{row}")
    
    row += 1
    # Box 1: Source Portfolio
    ws[f"A{row}"] = "SOURCE PORTFOLIO"
    apply_style(ws[f"A{row}"], styles["header_sub"])
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row+1}"] = "Active Sources"
    ws[f"B{row+1}"] = "=Monthly_Data_Entry!B10"
    ws[f"A{row+2}"] = "Avg Quality"
    ws[f"B{row+2}"] = "=Monthly_Data_Entry!B11"
    ws[f"A{row+3}"] = "Coverage Gaps"
    ws[f"B{row+3}"] = "=Monthly_Data_Entry!B12"
    
    # Box 2: Production Quality
    ws[f"D{row}"] = "PRODUCTION QUALITY"
    apply_style(ws[f"D{row}"], styles["header_sub"])
    ws.merge_cells(f"D{row}:E{row}")
    ws[f"D{row+1}"] = "Products"
    ws[f"E{row+1}"] = "=Monthly_Data_Entry!B17"
    ws[f"D{row+2}"] = "Satisfaction"
    ws[f"E{row+2}"] = "=Monthly_Data_Entry!B19"
    ws[f"D{row+3}"] = "Consumption %"
    ws[f"E{row+3}"] = "=Monthly_Data_Entry!B20"
    
    # Box 3: Integration
    ws[f"G{row}"] = "INTEGRATION"
    apply_style(ws[f"G{row}"], styles["header_sub"])
    ws.merge_cells(f"G{row}:H{row}")
    ws[f"G{row+1}"] = "Tools Integrated"
    ws[f"H{row+1}"] = "=Monthly_Data_Entry!B25"
    ws[f"G{row+2}"] = "IOC Hit Rate %"
    ws[f"H{row+2}"] = "=Monthly_Data_Entry!B27"
    ws[f"G{row+3}"] = "False Positive %"
    ws[f"H{row+3}"] = "=Monthly_Data_Entry!B28"
    
    # Box 4: CVSS & Prevention
    ws[f"I{row}"] = "CVSS & PREVENTION"
    apply_style(ws[f"I{row}"], styles["header_sub"])
    ws.merge_cells(f"I{row}:J{row}")
    ws[f"I{row+1}"] = "CVSS 4.0 Adoption %"
    ws[f"J{row+1}"] = "=Monthly_Data_Entry!B32"
    ws[f"I{row+2}"] = "Incidents Prevented"
    ws[f"J{row+2}"] = "=Monthly_Data_Entry!B40"
    ws[f"I{row+3}"] = "High CVSS Open"
    ws[f"J{row+3}"] = "=Monthly_Data_Entry!B34"
    
    # Critical Metrics Table
    row += 5
    ws[f"A{row}"] = "CRITICAL METRICS - THIS MONTH VS TARGET"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:J{row}")
    
    row += 1
    metric_headers = ["Metric", "Target", "Last Month", "This Month", "Status", "Trend"]
    for col_idx, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    row += 1
    critical_metrics = [
        ("Stakeholder Satisfaction", ">=4.0", "=OFFSET(Trend_History!F:F,-1,0)", "=Monthly_Data_Entry!B19", "", ""),
        ("IOC Hit Rate %", ">=10", "", "=Monthly_Data_Entry!B27", "", ""),
        ("Incidents Prevented", ">=3", "", "=Monthly_Data_Entry!B40", "", ""),
        ("CVSS Accuracy %", ">=90", "", "=Monthly_Data_Entry!B33", "", ""),
        ("False Positive %", "<=5", "", "=Monthly_Data_Entry!B28", "", ""),
    ]
    
    for metric_data in critical_metrics:
        for col_idx, value in enumerate(metric_data, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        row += 1
    
    # Top Items Section (Bottom)
    row += 1
    ws[f"A{row}"] = "TOP 3 ACHIEVEMENTS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:C{row}")
    
    ws[f"D{row}"] = "TOP 3 RISKS"
    apply_style(ws[f"D{row}"], styles["section_header"])
    ws.merge_cells(f"D{row}:G{row}")
    
    ws[f"H{row}"] = "TOP 3 ACTIONS REQUIRED"
    apply_style(ws[f"H{row}"], styles["section_header"])
    ws.merge_cells(f"H{row}:J{row}")
    
    row += 1
    # Placeholder for dynamic content (would use formulas to pull from other sheets)
    ws[f"A{row}"] = "• [From Critical_Actions, Type=Achievement]"
    ws[f"D{row}"] = "• [From Risk_Summary, Priority=Critical]"
    ws[f"H{row}"] = "• [From Critical_Actions, Priority=High]"
    
    # Footer
    row += 5
    ws[f"A{row}"] = f"Data Source: Self-reported metrics | Next Update: 1st business day of next month"
    ws.merge_cells(f"A{row}:J{row}")
    
    # Column widths
    set_column_widths(ws, [15, 12, 12, 12, 12, 12, 12, 12, 12, 12])
    
    logger.info("✅ Sheet 3: Executive_Dashboard - Complete")


# ============================================================================
# SECTION 5: SHEET 4 - TREND HISTORY
# ============================================================================

def create_trend_history_sheet(wb):
    """Create Sheet 4: Trend_History - 12-month historical data."""
    ws = wb["Trend_History"]
    styles = get_styles()
    
    # Title
    ws["A1"] = "Trend History - 12 Month Rolling Data"
    apply_style(ws["A1"], styles["header_main"])
    ws.merge_cells("A1:Q1")
    
    ws["A2"] = "Data automatically appended from Monthly_Data_Entry when Status = 'Final'"
    apply_style(ws["A2"], styles["header_sub"])
    ws.merge_cells("A2:Q2")
    
    # Column headers
    row = 4
    headers = [
        "Period", "Entry_Date", "Active_Sources", "Avg_Quality", "Coverage_Gaps",
        "Products_Published", "Satisfaction", "IOCs_Deployed", "Hit_Rate_%",
        "False_Positive_%", "CVSS_4.0_%", "CVSS_Accuracy_%", "High_CVSS_Open",
        "Prevented", "Cost_Avoid_CHF", "Program_Health", "Notes"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    # Add example row
    row += 1
    example_data = [
        "2025-01", "2025-02-01", 18, 4.2, 2, 23, 4.3, 1247, 12.3,
        3.8, 68, 91, 0, 4, 275000, 4.1, "Q1 data"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        ws.cell(row=row, column=col_idx, value=value)
    
    # Instructions
    row += 2
    ws[f"A{row}"] = "INSTRUCTIONS:"
    apply_style(ws[f"A{row}"], styles["section_header"])
    
    row += 1
    ws[f"A{row}"] = "• This sheet stores last 24 months of data for trending"
    row += 1
    ws[f"A{row}"] = "• New row appended automatically when Monthly_Data_Entry Status = 'Final'"
    row += 1
    ws[f"A{row}"] = "• Used for sparklines and charts in Executive_Dashboard"
    row += 1
    ws[f"A{row}"] = "• Keep last 24 months, delete older rows"
    
    # Column widths
    widths = [12, 12, 12, 10, 12, 12, 10, 12, 10, 12, 10, 12, 12, 10, 14, 12, 20]
    set_column_widths(ws, widths)
    
    logger.info("✅ Sheet 4: Trend_History - Complete")


# ============================================================================
# SECTION 6: SHEET 5 - CRITICAL ACTIONS
# ============================================================================

def create_critical_actions_sheet(wb):
    """Create Sheet 5: Critical_Actions - Top executive actions."""
    ws = wb["Critical_Actions"]
    styles = get_styles()
    
    # Title
    ws["A1"] = "Critical Actions - Executive Attention Required"
    apply_style(ws["A1"], styles["header_main"])
    ws.merge_cells("A1:J1")
    
    ws["A2"] = "Track 5-10 actions requiring executive decisions or awareness"
    apply_style(ws["A2"], styles["header_sub"])
    ws.merge_cells("A2:J2")
    
    # Column headers
    row = 4
    headers = [
        "Action_ID", "Action_Type", "Description", "Owner", "Priority",
        "Due_Date", "Status", "Resolution", "Created_Date", "Completed_Date"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    # Add data validation
    # Action_Type dropdown
    action_type_dv = DataValidation(
        type="list",
        formula1='"Achievement,Risk_Mitigation,Budget,Resource,Strategic"',
        allow_blank=False
    )
    ws.add_data_validation(action_type_dv)
    action_type_dv.add(f"B5:B100")
    
    # Priority dropdown
    priority_dv = DataValidation(
        type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚫ Low"',
        allow_blank=False
    )
    ws.add_data_validation(priority_dv)
    priority_dv.add(f"E5:E100")
    
    # Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Open,In_Progress,Completed,Cancelled"',
        allow_blank=False
    )
    ws.add_data_validation(status_dv)
    status_dv.add(f"G5:G100")
    
    # Example rows
    row += 1
    examples = [
        ("ACT-2025-01", "Budget", "CrowdStrike renewal requires board approval", "CISO", "High", 
         "31.03.2025", "Open", "", "10.01.2025", ""),
        ("ACT-2025-02", "Achievement", "Prevented 4 major incidents using threat intelligence", "TI Lead", "Medium",
         "31.01.2025", "Completed", "Documented in monthly report", "05.01.2025", "31.01.2025"),
    ]
    
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        row += 1
    
    # Summary section
    row += 2
    ws[f"A{row}"] = "SUMMARY METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:D{row}")
    
    row += 1
    ws[f"A{row}"] = "Total Actions:"
    ws[f"B{row}"] = f"=COUNTA(A5:A100)"
    ws[f"A{row+1}"] = "Critical Priority:"
    ws[f"B{row+1}"] = f'=COUNTIF(E5:E100,"Critical")'
    ws[f"A{row+2}"] = "Overdue:"
    ws[f"B{row+2}"] = f'=COUNTIFS(F5:F100,"<"&TODAY(),G5:G100,"<>Completed")'
    ws[f"A{row+3}"] = "Completed This Quarter:"
    ws[f"B{row+3}"] = f'=COUNTIFS(J5:J100,">="&DATE(YEAR(TODAY()),QUARTER(TODAY())*3-2,1),G5:G100,"Completed")'
    
    # Column widths
    set_column_widths(ws, [12, 18, 40, 15, 12, 12, 15, 30, 14, 14])
    
    logger.info("✅ Sheet 5: Critical_Actions - Complete")


# This is part 2 - continue with part 3...
# ============================================================================
# SECTION 7: SHEET 6 - QUARTERLY SUMMARY
# ============================================================================

def create_quarterly_summary_sheet(wb):
    """Create Sheet 6: Quarterly_Summary - Quarter-over-quarter comparison."""
    ws = wb["Quarterly_Summary"]
    styles = get_styles()
    
    # Title
    ws["A1"] = "Quarterly Summary - Board Reporting"
    apply_style(ws["A1"], styles["header_main"])
    ws.merge_cells("A1:G1")
    
    ws["A2"] = "Quarter-over-quarter comparison and strategic highlights"
    apply_style(ws["A2"], styles["header_sub"])
    ws.merge_cells("A2:G2")
    
    # Section A: Quarterly Metrics
    row = 4
    ws[f"A{row}"] = "A. QUARTERLY METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:G{row}")
    
    row += 1
    headers = ["Metric", "Q-4", "Q-3", "Q-2", "Q-1 (Current)", "QoQ Change", "YoY Change"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    row += 1
    # Source Portfolio section
    ws[f"A{row}"] = "SOURCE PORTFOLIO"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1
    
    metrics = [
        ("Active Sources", "", "", "", "", "", ""),
        ("Avg Quality (1-5)", "", "", "", "", "", ""),
        ("Coverage Gaps", "", "", "", "", "", ""),
    ]
    
    for metric_data in metrics:
        for col_idx, value in enumerate(metric_data, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        row += 1
    
    # Production section
    ws[f"A{row}"] = "PRODUCTION"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1
    
    metrics = [
        ("Products (Quarterly)", "", "", "", "", "", ""),
        ("Satisfaction (1-5)", "", "", "", "", "", ""),
    ]
    
    for metric_data in metrics:
        for col_idx, value in enumerate(metric_data, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        row += 1
    
    # Integration section
    ws[f"A{row}"] = "INTEGRATION"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1
    
    metrics = [
        ("IOCs Deployed (Q)", "", "", "", "", "", ""),
        ("Hit Rate (%)", "", "", "", "", "", ""),
    ]
    
    for metric_data in metrics:
        for col_idx, value in enumerate(metric_data, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        row += 1
    
    # CVSS & Prevention section
    ws[f"A{row}"] = "CVSS & PREVENTION"
    apply_style(ws[f"A{row}"], styles["section_header"])
    row += 1
    
    metrics = [
        ("CVSS 4.0 Adoption (%)", "", "", "", "", "", ""),
        ("Incidents Prevented (Q)", "", "", "", "", "", ""),
        ("Cost Avoidance (CHF, Q)", "", "", "", "", "", ""),
    ]
    
    for metric_data in metrics:
        for col_idx, value in enumerate(metric_data, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        row += 1
    
    # Section B: Quarterly Highlights
    row += 2
    ws[f"A{row}"] = "B. QUARTERLY HIGHLIGHTS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:G{row}")
    
    row += 1
    ws[f"A{row}"] = "Quarter"
    ws[f"B{row}"] = "Top 3 Achievements"
    ws[f"E{row}"] = "Top 3 Challenges"
    apply_style(ws[f"A{row}"], styles["header_sub"])
    apply_style(ws[f"B{row}"], styles["header_sub"])
    apply_style(ws[f"E{row}"], styles["header_sub"])
    ws.merge_cells(f"B{row}:D{row}")
    ws.merge_cells(f"E{row}:G{row}")
    
    row += 1
    quarters = ["Q-1 (Current)", "Q-2", "Q-3"]
    for quarter in quarters:
        ws.cell(row=row, column=1, value=quarter)
        ws.merge_cells(f"B{row}:D{row+2}")
        ws.merge_cells(f"E{row}:G{row+2}")
        row += 3
    
    # Section C: Board Observations
    row += 1
    ws[f"A{row}"] = "C. CISO COMMENTARY - STRATEGIC OBSERVATIONS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:G{row}")
    
    row += 1
    ws.merge_cells(f"A{row}:G{row+5}")
    ws[f"A{row}"] = "[CISO Commentary: Quarterly strategic priorities, budget implications, resource needs - max 500 characters]"
    
    # Column widths
    set_column_widths(ws, [25, 12, 12, 12, 15, 13, 13])
    
    logger.info("✅ Sheet 6: Quarterly_Summary - Complete")


# ============================================================================
# SECTION 8: SHEET 7 - RISK SUMMARY
# ============================================================================

def create_risk_summary_sheet(wb):
    """Create Sheet 7: Risk_Summary - Top program risks."""
    ws = wb["Risk_Summary"]
    styles = get_styles()
    
    # Title
    ws["A1"] = "Risk Summary - Program Risks Requiring Executive Attention"
    apply_style(ws["A1"], styles["header_main"])
    ws.merge_cells("A1:K1")
    
    ws["A2"] = "Top 5-10 risks to threat intelligence program effectiveness"
    apply_style(ws["A2"], styles["header_sub"])
    ws.merge_cells("A2:K2")
    
    # Column headers
    row = 4
    headers = [
        "Risk_ID", "Risk_Category", "Risk_Description", "Likelihood",
        "Impact", "Risk_Score", "Mitigation_Status", "Mitigation_Plan",
        "Owner", "Review_Date", "Status"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    # Add data validation
    # Risk_Category dropdown
    category_dv = DataValidation(
        type="list",
        formula1='"Source_Portfolio,Collection,Integration,Staffing,Budget,Compliance,Technical"',
        allow_blank=False
    )
    ws.add_data_validation(category_dv)
    category_dv.add(f"B5:B100")
    
    # Likelihood dropdown
    likelihood_dv = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(likelihood_dv)
    likelihood_dv.add(f"D5:D100")
    
    # Impact dropdown
    impact_dv = DataValidation(
        type="list",
        formula1='"🔴 Critical,🟡 High,🟢 Medium,⚫ Low"',
        allow_blank=False
    )
    ws.add_data_validation(impact_dv)
    impact_dv.add(f"E5:E100")
    
    # Mitigation_Status dropdown
    mitigation_dv = DataValidation(
        type="list",
        formula1='"None,Planned,In_Progress,Mitigated"',
        allow_blank=False
    )
    ws.add_data_validation(mitigation_dv)
    mitigation_dv.add(f"G5:G100")
    
    # Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Active,Closed,Accepted"',
        allow_blank=False
    )
    ws.add_data_validation(status_dv)
    status_dv.add(f"K5:K100")
    
    # Example rows
    row += 1
    examples = [
        ("RISK-2025-01", "Staffing", "Lead threat analyst departure without trained backup",
         "High", "High", "9", "Planned", "Cross-training analyst 2", "CISO", "15.01.2025", "Active"),
        ("RISK-2025-02", "Source_Portfolio", "Over-reliance on single commercial vendor",
         "Medium", "High", "6", "In_Progress", "Diversifying to 3 vendors", "TI Lead", "20.01.2025", "Active"),
    ]
    
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        # Add Risk_Score formula
        ws.cell(row=row, column=6, value=f"=IF(D{row}=\"High\",IF(E{row}=\"Critical\",12,IF(E{row}=\"High\",9,IF(E{row}=\"Medium\",6,3))),IF(D{row}=\"Medium\",IF(E{row}=\"Critical\",8,IF(E{row}=\"High\",6,IF(E{row}=\"Medium\",4,2))),IF(E{row}=\"Critical\",4,IF(E{row}=\"High\",3,IF(E{row}=\"Medium\",2,1)))))")
        row += 1
    
    # Risk Score Matrix Reference
    row += 2
    ws[f"A{row}"] = "RISK SCORE MATRIX"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 1
    ws[f"A{row}"] = "Likelihood \\ Impact"
    ws[f"B{row}"] = "Low"
    ws[f"C{row}"] = "Medium"
    ws[f"D{row}"] = "High"
    ws[f"E{row}"] = "Critical"
    apply_style(ws[f"A{row}"], styles["header_sub"])
    
    row += 1
    matrix = [
        ("High", "3", "6", "9", "12"),
        ("Medium", "2", "4", "6", "8"),
        ("Low", "1", "2", "3", "4"),
    ]
    
    for matrix_row in matrix:
        for col_idx, value in enumerate(matrix_row, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        row += 1
    
    # Summary Metrics
    row += 1
    ws[f"A{row}"] = "SUMMARY METRICS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:D{row}")
    
    row += 1
    ws[f"A{row}"] = "Total Active Risks:"
    ws[f"B{row}"] = f'=COUNTIF(K5:K100,"Active")'
    ws[f"A{row+1}"] = "Critical Risks (Score ≥9):"
    ws[f"B{row+1}"] = f'=COUNTIFS(F5:F100,">=9",K5:K100,"Active")'
    ws[f"A{row+2}"] = "High Risks (Score 6-8):"
    ws[f"B{row+2}"] = f'=COUNTIFS(F5:F100,">=6",F5:F100,"<9",K5:K100,"Active")'
    ws[f"A{row+3}"] = "Risks Needing Mitigation:"
    ws[f"B{row+3}"] = f'=COUNTIFS(G5:G100,"None",K5:K100,"Active")'
    
    # Column widths
    set_column_widths(ws, [14, 18, 35, 12, 12, 10, 16, 30, 15, 12, 10])
    
    logger.info("✅ Sheet 7: Risk_Summary - Complete")


# ============================================================================
# SECTION 9: SHEET 8 - ROI SUMMARY
# ============================================================================

def create_roi_summary_sheet(wb):
    """Create Sheet 8: ROI_Summary - Cost-benefit analysis."""
    ws = wb["ROI_Summary"]
    styles = get_styles()
    
    # Title
    ws["A1"] = "ROI Summary - Threat Intelligence Program Value"
    apply_style(ws["A1"], styles["header_main"])
    ws.merge_cells("A1:F1")
    
    ws["A2"] = "Cost-benefit analysis and program value justification"
    apply_style(ws["A2"], styles["header_sub"])
    ws.merge_cells("A2:F2")
    
    # Section A: Annual Program Costs
    row = 4
    ws[f"A{row}"] = "A. ANNUAL PROGRAM COSTS"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 1
    headers = ["Cost Category", "This Year (CHF)", "Last Year (CHF)", "Change (CHF)", "Change (%)", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    row += 1
    cost_categories = [
        ("Commercial TI Sources", 450000, 420000, "=B{}-C{}", "=(B{}-C{})/C{}*100", ""),
        ("Staffing (FTE cost)", 600000, 580000, "=B{}-C{}", "=(B{}-C{})/C{}*100", ""),
        ("Tools & Infrastructure", 150000, 140000, "=B{}-C{}", "=(B{}-C{})/C{}*100", ""),
        ("Training & Certifications", 25000, 20000, "=B{}-C{}", "=(B{}-C{})/C{}*100", ""),
        ("Travel & Conferences", 15000, 12000, "=B{}-C{}", "=(B{}-C{})/C{}*100", ""),
    ]
    
    start_row = row
    for category, this_year, last_year, change_formula, pct_formula, notes in cost_categories:
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=2, value=this_year)
        ws.cell(row=row, column=3, value=last_year)
        ws.cell(row=row, column=4, value=change_formula.format(row, row))
        ws.cell(row=row, column=5, value=pct_formula.format(row, row, row))
        ws.cell(row=row, column=6, value=notes)
        row += 1
    
    # Total row
    ws[f"A{row}"] = "TOTAL PROGRAM COST"
    apply_style(ws[f"A{row}"], styles["header_sub"])
    ws[f"B{row}"] = f"=SUM(B{start_row}:B{row-1})"
    ws[f"C{row}"] = f"=SUM(C{start_row}:C{row-1})"
    ws[f"D{row}"] = f"=B{row}-C{row}"
    ws[f"E{row}"] = f"=(B{row}-C{row})/C{row}*100"
    
    # Section B: Value Delivered
    row += 2
    ws[f"A{row}"] = "B. VALUE DELIVERED"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 1
    headers = ["Value Category", "This Year (CHF)", "Last Year (CHF)", "Change (CHF)", "Change (%)", "Calculation Method"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    row += 1
    value_categories = [
        ("Incidents Prevented (Cost Avoidance)", 3300000, 2800000, "=B{}-C{}", "=(B{}-C{})/C{}*100", "From Trend_History"),
        ("Faster Incident Response (Time Savings)", 450000, 380000, "=B{}-C{}", "=(B{}-C{})/C{}*100", "120 hrs × 3750/hr"),
        ("Risk Assessment Improvements", 180000, 150000, "=B{}-C{}", "=(B{}-C{})/C{}*100", "12 updates × 15K"),
        ("Compliance & Audit Efficiency", 90000, 75000, "=B{}-C{}", "=(B{}-C{})/C{}*100", "60 hrs × 1500/hr"),
    ]
    
    start_row_value = row
    for category, this_year, last_year, change_formula, pct_formula, method in value_categories:
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=2, value=this_year)
        ws.cell(row=row, column=3, value=last_year)
        ws.cell(row=row, column=4, value=change_formula.format(row, row))
        ws.cell(row=row, column=5, value=pct_formula.format(row, row, row))
        ws.cell(row=row, column=6, value=method)
        row += 1
    
    # Total value row
    ws[f"A{row}"] = "TOTAL VALUE DELIVERED"
    apply_style(ws[f"A{row}"], styles["header_sub"])
    ws[f"B{row}"] = f"=SUM(B{start_row_value}:B{row-1})"
    ws[f"C{row}"] = f"=SUM(C{start_row_value}:C{row-1})"
    ws[f"D{row}"] = f"=B{row}-C{row}"
    ws[f"E{row}"] = f"=(B{row}-C{row})/C{row}*100"
    
    # Section C: ROI Calculation
    row += 2
    ws[f"A{row}"] = "C. ROI CALCULATION"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 1
    headers = ["Metric", "This Year", "Last Year", "3-Year Avg", "Interpretation", ""]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    row += 1
    # Reference the totals from above - will need to adjust formulas
    total_cost_row = start_row + 5  # Approximate
    total_value_row = start_row_value + 4  # Approximate
    
    roi_metrics = [
        ("Total Value Delivered (CHF)", f"=B{total_value_row}", f"=C{total_value_row}", "", "", ""),
        ("Total Program Cost (CHF)", f"=B{total_cost_row}", f"=C{total_cost_row}", "", "", ""),
        ("Net Value (CHF)", f"=B{row-2}-B{row-1}", f"=C{row-2}-C{row-1}", "", "", ""),
        ("ROI Ratio", f"=B{row-3}/B{row-2}", f"=C{row-3}/C{row-2}", "", ">=3.0 Excellent", ""),
        ("ROI Percentage", f"=(B{row-3}/B{row-2}-1)*100", f"=(C{row-3}/C{row-2}-1)*100", "", ">=200% Excellent", ""),
    ]
    
    for metric_data in roi_metrics:
        for col_idx, value in enumerate(metric_data, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        row += 1
    
    # Section D: CVSS-Based Prevention Value
    row += 2
    ws[f"A{row}"] = "D. CVSS-BASED PREVENTION VALUE (Annual)"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:F{row}")
    
    row += 1
    headers = ["CVSS Severity", "Incidents Prevented", "Avg Cost/Incident", "Total Value (CHF)", "% of Total", ""]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, styles["header_sub"])
    
    row += 1
    cvss_data = [
        ("Critical (9.0-10.0)", 8, 150000, "=B{}*C{}", "=D{}/D{}", ""),
        ("High (7.0-8.9)", 18, 75000, "=B{}*C{}", "=D{}/D{}", ""),
        ("Medium (4.0-6.9)", 14, 25000, "=B{}*C{}", "=D{}/D{}", ""),
        ("Low (0.1-3.9)", 4, 5000, "=B{}*C{}", "=D{}/D{}", ""),
    ]
    
    start_cvss = row
    for severity, count, avg_cost, total_formula, pct_formula, notes in cvss_data:
        ws.cell(row=row, column=1, value=severity)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=avg_cost)
        ws.cell(row=row, column=4, value=total_formula.format(row, row))
        row += 1
    
    # Total CVSS row
    ws[f"A{row}"] = "TOTAL"
    apply_style(ws[f"A{row}"], styles["header_sub"])
    ws[f"B{row}"] = f"=SUM(B{start_cvss}:B{row-1})"
    ws[f"C{row}"] = f"=AVERAGE(C{start_cvss}:C{row-1})"
    ws[f"D{row}"] = f"=SUM(D{start_cvss}:D{row-1})"
    
    # Add percentages
    for r in range(start_cvss, row):
        ws[f"E{r}"] = f"=D{r}/D${row}*100"
    
    # Column widths
    set_column_widths(ws, [35, 15, 15, 15, 12, 25])
    
    logger.info("✅ Sheet 8: ROI_Summary - Complete")


# This is part 3 - continue with part 4 (Metadata and main)...
# ============================================================================
# SECTION 10: SHEET 9 - METADATA
# ============================================================================

def create_metadata_sheet(wb):
    """Create Sheet 9: Metadata - Document control."""
    ws = wb["Metadata"]
    styles = get_styles()
    
    # Title
    ws["A1"] = "Workbook Metadata - Document Control"
    apply_style(ws["A1"], styles["header_main"])
    ws.merge_cells("A1:C1")
    
    # Metadata fields
    row = 3
    metadata = [
        ("Workbook_Version", "1.0"),
        ("Workbook_Type", "Standalone Dashboard (Manual Entry)"),
        ("Total_Sheets", "9"),
        ("Generation_Date", datetime.now().strftime("%d.%m.%Y")),
        ("Generator_Script", "generate_a57_5_standalone.py"),
        ("Script_Version", "1.0.0"),
        ("Python_Version", "3.x"),
        ("openpyxl_Version", "3.x"),
        ("Last_Modified_Date", "=TODAY()"),
        ("Last_Modified_By", "=USER()"),
        ("Reporting_Period_Current", "=Monthly_Data_Entry!B5"),
        ("Data_Entry_Status", "=Monthly_Data_Entry!B8"),
        ("Related_Policy", "ISMS-POL-A.5.7"),
        ("Related_IMP_Spec", "ISMS-IMP-A.5.7.5 v1.0"),
        ("Complementary_Dashboard", "ISMS-IMP-A.5.7.4 (Comprehensive)"),
        ("ISO_27001_Control", "A.5.7 (Threat Intelligence)"),
        ("Audience", "C-suite, Board, External Stakeholders"),
        ("Update_Frequency", "Monthly (1st business day)"),
        ("Quarterly_Review", "Last week of quarter"),
    ]
    
    for field, value in metadata:
        ws.cell(row=row, column=1, value=field)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    # Changelog
    row += 2
    ws[f"A{row}"] = "CHANGELOG"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:C{row}")
    
    row += 1
    changelog = """v1.0 (10.01.2025):
- Initial release with 9 sheets
- Monthly manual data entry workflow (10 minutes)
- Executive one-page dashboard (auto-generated)
- 12-month trend tracking
- Quarterly Board reporting
- Top risks and critical actions tracking
- ROI calculation and cost-benefit analysis
- CVSS-based prevention value tracking
- ISO 27001 compliance status
- Completely standalone (no external workbook dependencies)
"""
    
    ws[f"A{row}"] = changelog
    ws.merge_cells(f"A{row}:C{row+10}")
    ws.row_dimensions[row].height = 150
    
    # Sheet inventory
    row += 12
    ws[f"A{row}"] = "SHEET INVENTORY"
    apply_style(ws[f"A{row}"], styles["section_header"])
    ws.merge_cells(f"A{row}:C{row}")
    
    row += 1
    sheets_info = [
        ("Sheet 1", "Instructions", "User guide and workflow"),
        ("Sheet 2", "Monthly_Data_Entry", "Manual entry form (15-20 metrics)"),
        ("Sheet 3", "Executive_Dashboard", "One-page visual summary (print-ready)"),
        ("Sheet 4", "Trend_History", "12-month historical data"),
        ("Sheet 5", "Critical_Actions", "Top 5-10 executive actions"),
        ("Sheet 6", "Quarterly_Summary", "QoQ comparison for Board"),
        ("Sheet 7", "Risk_Summary", "Top program risks"),
        ("Sheet 8", "ROI_Summary", "Cost-benefit analysis"),
        ("Sheet 9", "Metadata", "This sheet - document control"),
    ]
    
    for sheet_num, sheet_name, description in sheets_info:
        ws.cell(row=row, column=1, value=sheet_num)
        ws.cell(row=row, column=2, value=sheet_name)
        ws.cell(row=row, column=3, value=description)
        row += 1
    
    # Column widths
    set_column_widths(ws, [30, 35, 50])
    
    logger.info("✅ Sheet 9: Metadata - Complete")


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function - generates the complete workbook."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.5.7.5 - Threat Intelligence Standalone Dashboard Generator")
    logger.info("ISO/IEC 27001:2022 Control A.5.7 (Threat Intelligence)")
    logger.info("=" * 80)
    logger.info("")
    
    # Create workbook
    logger.info("Creating workbook structure...")
    wb = create_workbook()
    logger.info("✅ Workbook created with 9 sheets")
    logger.info("")
    
    # Generate all sheets
    logger.info("Generating sheets:")
    logger.info("-" * 80)
    
    create_instructions_sheet(wb)
    create_monthly_data_entry_sheet(wb)
    create_executive_dashboard_sheet(wb)
    create_trend_history_sheet(wb)
    create_critical_actions_sheet(wb)
    create_quarterly_summary_sheet(wb)
    create_risk_summary_sheet(wb)
    create_roi_summary_sheet(wb)
    create_metadata_sheet(wb)
    
    logger.info("-" * 80)
    logger.info("")
    
    # Save workbook
    filename = f"ISMS-IMP-A.5.7.5_Standalone_Dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    logger.info("=" * 80)
    logger.info("✅ WORKBOOK GENERATED SUCCESSFULLY")
    logger.info("=" * 80)
    logger.info("")
    logger.info(f"Generated: {filename}")
    logger.info("")
    logger.info("Workbook Contents:")
    logger.info("  • 9 sheets: Instructions through Metadata")
    logger.info("  • Manual entry form with 15-20 key TI metrics")
    logger.info("  • Auto-generated executive one-page dashboard (print-ready)")
    logger.info("  • 12-month trend tracking")
    logger.info("  • Quarterly Board reporting artifacts")
    logger.info("  • Top risks and critical actions tracking")
    logger.info("  • ROI calculation with CVSS-based prevention value")
    logger.info("  • ISO 27001 Control A.5.7 compliance status")
    logger.info("  • Completely STANDALONE (no external workbook dependencies)")
    logger.info("")
    logger.info("Key Features:")
    logger.info("  ✅ 10-minute monthly update workflow")
    logger.info("  ✅ Executive-focused, not operational")
    logger.info("  ✅ Board-ready quarterly reports")
    logger.info("  ✅ Shareable without dependencies")
    logger.info("  ✅ Strategic visibility without operational detail exposure")
    logger.info("")
    logger.info("=" * 80)
    logger.info("📋 NEXT STEPS")
    logger.info("=" * 80)
    logger.info("")
    logger.info("1. MONTHLY WORKFLOW (10 minutes):")
    logger.info("   • Open workbook on 1st business day of new month")
    logger.info("   • Navigate to 'Monthly_Data_Entry' sheet")
    logger.info("   • Fill in 15-20 key metrics")
    logger.info("   • Set Status to 'Final'")
    logger.info("   • Review auto-generated 'Executive_Dashboard'")
    logger.info("   • Export Dashboard to PDF for distribution")
    logger.info("")
    logger.info("2. QUARTERLY WORKFLOW (additional 30 minutes):")
    logger.info("   • Complete monthly workflow")
    logger.info("   • Update 'Quarterly_Summary' sheet")
    logger.info("   • Update 'ROI_Summary' sheet")
    logger.info("   • Prepare Board presentation (3 PDFs)")
    logger.info("")
    logger.info("3. DATA SOURCES:")
    logger.info("   • Active Sources: ISMS-IMP-A.5.7.1 or TI platform")
    logger.info("   • Production Metrics: Document repository")
    logger.info("   • IOC Data: SIEM dashboard")
    logger.info("   • CVSS Metrics: Vulnerability management platform")
    logger.info("   • Prevention Data: ISMS-IMP-A.5.7.3")
    logger.info("")
    logger.info("4. VALIDATION:")
    logger.info("   • Run: python3 excel_sanity_check_a57_5.py <filename>")
    logger.info("   • Verify all 9 sheets present")
    logger.info("   • Verify formulas not broken")
    logger.info("   • Test data entry workflow")
    logger.info("")
    logger.info("5. DISTRIBUTION:")
    logger.info("   • Store in secure location (SharePoint, encrypted drive)")
    logger.info("   • Set appropriate file permissions")
    logger.info("   • Schedule monthly reminder (1st business day)")
    logger.info("   • Archive historical copies (keep 3 years)")
    logger.info("")
    logger.info("COMPLEMENTARY DASHBOARDS:")
    logger.info("  • ISMS-IMP-A.5.7.4 (Comprehensive) - For internal operations")
    logger.info("  • ISMS-IMP-A.5.7.5 (This) - For executive/external reporting")
    logger.info("")
    logger.info("DIFFERENCE:")
    logger.info("  • 5.7.4 = Automated, external references, operational detail")
    logger.info("  • 5.7.5 = Manual entry, standalone, executive summary")
    logger.info("")
    logger.info("=" * 80)
    logger.info("For questions or issues:")
    logger.info("  • Documentation: See 'Instructions' sheet in workbook")
    logger.info("  • Policy: ISMS-POL-A.5.7 (Threat Intelligence Policy)")
    logger.info("  • Specification: ISMS-IMP-A.5.7.5 (Implementation Spec)")
    logger.info("=" * 80)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
