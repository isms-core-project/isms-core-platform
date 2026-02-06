#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.35-36.S4 - Compliance & Review Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.5.35-36: Compliance & Review

Assessment Domain 4 of 4: Compliance Dashboard

This script generates a comprehensive dashboard for monitoring overall ISMS
compliance and review status, consolidating metrics from independent reviews
and compliance assessments.

**Generated Workbook Structure:**
1. Instructions - Dashboard guidance
2. Executive_Summary - High-level compliance overview
3. Review_Status - Independent review tracking
4. Compliance_Status - Policy compliance overview
5. Findings_Overview - Findings metrics and aging
6. KPI_Scorecard - Key performance indicators
7. Trend_Analysis - Historical compliance trends
8. Approval_SignOff - Management approval

================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

DOCUMENT_ID = "ISMS-IMP-A.5.35-36.S4"
WORKBOOK_NAME = "Compliance and Review Dashboard"
CONTROL_ID = "A.5.35-36"
CONTROL_NAME = "Compliance & Review"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


def setup_styles():
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "kpi_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "kpi_value": {
            "font": Font(name="Calibri", size=18, bold=True),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
        "green": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "amber": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "red": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }


def create_workbook() -> Workbook:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    for name in ["Instructions", "Executive_Summary", "Review_Status",
                 "Compliance_Status", "Findings_Overview", "KPI_Scorecard",
                 "Trend_Analysis", "Approval_SignOff"]:
        wb.create_sheet(title=name)
    return wb


def create_instructions_sheet(ws, styles):
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Dashboard Type", "ISMS Compliance & Review Dashboard"),
        ("Control Reference", CONTROL_ID),
        ("Version", "1.0"),
        ("Reporting Period", ""),
        ("Prepared By", ""),
        ("Organisation", ""),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "DASHBOARD PURPOSE"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    ws[f"A{row}"] = (
        "This dashboard provides executive visibility into ISMS compliance and review "
        "status, consolidating metrics from independent reviews (A.5.35) and compliance "
        "assessments (A.5.36) to support management decision-making."
    )
    ws.merge_cells(f"A{row}:G{row}")

    row += 2
    ws[f"A{row}"] = "DASHBOARD COMPONENTS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    components = [
        ("Executive Summary", "High-level KPIs and overall compliance status"),
        ("Review Status", "Independent review schedule and completion tracking"),
        ("Compliance Status", "Policy and control compliance overview"),
        ("Findings Overview", "Findings metrics, aging, and severity breakdown"),
        ("KPI Scorecard", "Detailed key performance indicators"),
        ("Trend Analysis", "Historical compliance trends and patterns"),
    ]

    row += 1
    for component, desc in components:
        ws[f"A{row}"] = component
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = desc
        row += 1

    row += 1
    ws[f"A{row}"] = "RAG STATUS DEFINITIONS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    rag = [
        ("Green", "On track; no significant issues; targets being met"),
        ("Amber", "At risk; minor issues; attention needed"),
        ("Red", "Off track; significant issues; immediate action required"),
    ]

    row += 1
    for status, desc in rag:
        ws[f"A{row}"] = status
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = desc
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 55
    ws.freeze_panes = "A4"


def create_executive_summary_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "EXECUTIVE SUMMARY - ISMS Compliance & Review"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # KPI Boxes Row 1
    # Overall Compliance
    ws.merge_cells("A3:B3")
    ws["A3"] = "OVERALL COMPLIANCE"
    ws["A3"].font = styles["kpi_header"]["font"]
    ws["A3"].fill = styles["kpi_header"]["fill"]
    ws["A3"].alignment = styles["kpi_header"]["alignment"]

    ws.merge_cells("A4:B5")
    ws["A4"].font = styles["kpi_value"]["font"]
    ws["A4"].alignment = styles["kpi_value"]["alignment"]
    ws["A4"].fill = styles["input_cell"]["fill"]
    ws["A4"].border = styles["border"]

    # Review Completion
    ws.merge_cells("D3:E3")
    ws["D3"] = "REVIEW COMPLETION"
    ws["D3"].font = styles["kpi_header"]["font"]
    ws["D3"].fill = styles["kpi_header"]["fill"]
    ws["D3"].alignment = styles["kpi_header"]["alignment"]

    ws.merge_cells("D4:E5")
    ws["D4"].font = styles["kpi_value"]["font"]
    ws["D4"].alignment = styles["kpi_value"]["alignment"]
    ws["D4"].fill = styles["input_cell"]["fill"]
    ws["D4"].border = styles["border"]

    # Open Findings
    ws.merge_cells("G3:H3")
    ws["G3"] = "OPEN FINDINGS"
    ws["G3"].font = styles["kpi_header"]["font"]
    ws["G3"].fill = styles["kpi_header"]["fill"]
    ws["G3"].alignment = styles["kpi_header"]["alignment"]

    ws.merge_cells("G4:H5")
    ws["G4"].font = styles["kpi_value"]["font"]
    ws["G4"].alignment = styles["kpi_value"]["alignment"]
    ws["G4"].fill = styles["input_cell"]["fill"]
    ws["G4"].border = styles["border"]

    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 30

    # Key Metrics Table
    row = 8
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "KEY METRICS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    metrics = [
        ("Independent Reviews Completed (YTD)", "", "2", "On Track"),
        ("Independent Reviews Scheduled (YTD)", "", "4", ""),
        ("Compliance Assessments Completed (YTD)", "", "4", "On Track"),
        ("Policy Compliance Rate", "", "95%", "On Track"),
        ("Control Implementation Rate", "", "90%", "At Risk"),
        ("Open Major Non-Conformities", "", "0", "On Track"),
        ("Open Minor Non-Conformities", "", "3", "At Risk"),
        ("Overdue Remediations", "", "0", "On Track"),
        ("Average Days to Close Findings", "", "30", "On Track"),
        ("Certification Readiness", "", "85%", "On Track"),
    ]

    headers = ["Metric", "Current Value", "Target", "Status"]
    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    dv_status = DataValidation(
        type="list",
        formula1='"On Track,At Risk,Off Track"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    row += 1
    for metric, value, target, status in metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=3, value=target)
        ws.cell(row=row, column=4, value=status if status else "")
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = styles["border"]
        dv_status.add(ws.cell(row=row, column=4))
        row += 1

    # Key Observations
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "KEY OBSERVATIONS & ACTIONS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    headers2 = ["Observation/Action", "", "", "Priority", "Owner", "Due Date"]
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = headers2[0]
    ws[f"D{row}"] = headers2[3]
    ws[f"E{row}"] = headers2[4]
    ws[f"F{row}"] = headers2[5]
    for col in ["A", "D", "E", "F"]:
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws[f"{col}{row}"].border = styles["border"]

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_priority)

    for r in range(row + 1, row + 9):
        ws.merge_cells(f"A{r}:C{r}")
        ws[f"A{r}"].fill = styles["input_cell"]["fill"]
        ws[f"A{r}"].border = styles["border"]
        for col in ["D", "E", "F"]:
            ws[f"{col}{r}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{r}"].border = styles["border"]
        dv_priority.add(ws[f"D{r}"])

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14


def create_review_status_sheet(ws, styles):
    ws.merge_cells("A1:J1")
    ws["A1"] = "INDEPENDENT REVIEW STATUS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Review_ID", 16),
        ("Review_Type", 22),
        ("Scope", 35),
        ("Planned_Date", 14),
        ("Actual_Date", 14),
        ("Status", 16),
        ("Lead_Reviewer", 22),
        ("Findings_Count", 14),
        ("Report_Status", 16),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Annual Comprehensive,Surveillance,Ad-Hoc,Pre-Certification,Post-Incident"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_status = DataValidation(
        type="list",
        formula1='"Planned,In Progress,Completed,Postponed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_report = DataValidation(
        type="list",
        formula1='"Not Started,Draft,Final,Issued"',
        allow_blank=False
    )
    ws.add_data_validation(dv_report)

    for r in range(4, 24):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=2))
        dv_status.add(ws.cell(row=r, column=6))
        dv_report.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "C4"


def create_compliance_status_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "POLICY & CONTROL COMPLIANCE STATUS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Area", 30),
        ("Policies_Total", 14),
        ("Policies_Compliant", 16),
        ("Compliance_%", 14),
        ("Controls_Total", 14),
        ("Controls_Implemented", 18),
        ("Implementation_%", 16),
        ("RAG_Status", 14),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    areas = [
        "Organisational Controls (A.5)",
        "People Controls (A.6)",
        "Physical Controls (A.7)",
        "Technological Controls (A.8)",
        "TOTAL",
    ]

    dv_rag = DataValidation(
        type="list",
        formula1='"Green,Amber,Red"',
        allow_blank=False
    )
    ws.add_data_validation(dv_rag)

    for row_idx, area in enumerate(areas, start=4):
        ws.cell(row=row_idx, column=1, value=area)
        if area == "TOTAL":
            ws.cell(row=row_idx, column=1).font = Font(bold=True)

        for c in range(2, 9):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Compliance % formula
        ws.cell(row=row_idx, column=4).value = f'=IF(B{row_idx}>0,C{row_idx}/B{row_idx}*100,"")'
        ws.cell(row=row_idx, column=4).number_format = '0.0"%"'
        ws.cell(row=row_idx, column=4).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Implementation % formula
        ws.cell(row=row_idx, column=7).value = f'=IF(E{row_idx}>0,F{row_idx}/E{row_idx}*100,"")'
        ws.cell(row=row_idx, column=7).number_format = '0.0"%"'
        ws.cell(row=row_idx, column=7).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        dv_rag.add(ws.cell(row=row_idx, column=8))

    # TOTAL row formulas
    total_row = len(areas) + 3
    ws.cell(row=total_row, column=2).value = f"=SUM(B4:B{total_row-1})"
    ws.cell(row=total_row, column=3).value = f"=SUM(C4:C{total_row-1})"
    ws.cell(row=total_row, column=5).value = f"=SUM(E4:E{total_row-1})"
    ws.cell(row=total_row, column=6).value = f"=SUM(F4:F{total_row-1})"

    ws.freeze_panes = "B4"


def create_findings_overview_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "FINDINGS OVERVIEW"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Summary section
    row = 3
    ws[f"A{row}"] = "Findings Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_items = [
        ("Total Open Findings", ""),
        ("Major Non-Conformities (Open)", ""),
        ("Minor Non-Conformities (Open)", ""),
        ("Observations (Open)", ""),
        ("Findings Closed (YTD)", ""),
        ("Findings Opened (YTD)", ""),
        ("Average Age of Open Findings (Days)", ""),
        ("Overdue Findings (>30 days)", ""),
    ]

    row += 1
    for item, value in summary_items:
        ws[f"A{row}"] = item
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Aging Analysis
    row += 1
    ws[f"A{row}"] = "Findings Aging Analysis"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    aging_headers = ["Age Bucket", "Major NCs", "Minor NCs", "Observations", "Total"]
    row += 1
    for col_idx, header in enumerate(aging_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    aging_buckets = ["0-7 days", "8-14 days", "15-30 days", "31-60 days", "61-90 days", ">90 days"]
    row += 1
    for bucket in aging_buckets:
        ws.cell(row=row, column=1, value=bucket)
        for c in range(2, 6):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        # Total formula
        ws.cell(row=row, column=5).value = f"=SUM(B{row}:D{row})"
        ws.cell(row=row, column=5).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14


def create_kpi_scorecard_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "KPI SCORECARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("KPI_ID", 12),
        ("KPI_Name", 40),
        ("Category", 20),
        ("Current_Value", 16),
        ("Target", 14),
        ("Threshold_Amber", 16),
        ("Threshold_Red", 14),
        ("Status", 14),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    kpis = [
        ("KPI-01", "Independent Reviews Completed On Schedule", "A.5.35", "100%", "<90%", "<75%"),
        ("KPI-02", "Policy Compliance Rate", "A.5.36", "95%", "<90%", "<80%"),
        ("KPI-03", "Control Implementation Rate", "A.5.36", "90%", "<85%", "<75%"),
        ("KPI-04", "Open Major NCs", "A.5.35", "0", ">0", ">2"),
        ("KPI-05", "Open Minor NCs", "A.5.36", "<5", ">5", ">10"),
        ("KPI-06", "Findings Closure Rate (30 days)", "A.5.35", "90%", "<80%", "<60%"),
        ("KPI-07", "Overdue Remediations", "A.5.35", "0", ">0", ">3"),
        ("KPI-08", "Department Compliance Attestations", "A.5.36", "100%", "<95%", "<85%"),
        ("KPI-09", "Repeat Findings Rate", "A.5.35", "<5%", ">5%", ">15%"),
        ("KPI-10", "Certification Readiness Score", "Both", "85%", "<80%", "<70%"),
    ]

    dv_status = DataValidation(
        type="list",
        formula1='"Green,Amber,Red"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for row_idx, (kpi_id, name, category, target, amber, red) in enumerate(kpis, start=4):
        ws.cell(row=row_idx, column=1, value=kpi_id)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=category)
        ws.cell(row=row_idx, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row_idx, column=5, value=target)
        ws.cell(row=row_idx, column=6, value=amber)
        ws.cell(row=row_idx, column=7, value=red)
        ws.cell(row=row_idx, column=8).fill = styles["input_cell"]["fill"]
        for c in range(1, 9):
            ws.cell(row=row_idx, column=c).border = styles["border"]
        dv_status.add(ws.cell(row=row_idx, column=8))

    ws.freeze_panes = "C4"


def create_trend_analysis_sheet(ws, styles):
    ws.merge_cells("A1:I1")
    ws["A1"] = "COMPLIANCE TREND ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Period", 14),
        ("Policy_Compliance_%", 18),
        ("Control_Implementation_%", 20),
        ("Reviews_Completed", 16),
        ("Open_Findings", 14),
        ("Closed_Findings", 14),
        ("Avg_Closure_Days", 16),
        ("Overall_Score", 14),
        ("Notes", 35),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    periods = ["Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025",
               "Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"]

    for row_idx, period in enumerate(periods, start=4):
        ws.cell(row=row_idx, column=1, value=period)
        for c in range(2, 10):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    ws.freeze_panes = "B4"


def create_approval_signoff_sheet(ws, styles):
    ws.merge_cells("A1:E1")
    ws["A1"] = "DASHBOARD APPROVAL"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Dashboard Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Dashboard Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Reporting Period", ""),
        ("Overall Compliance Status", ""),
        ("Review Completion Status", ""),
        ("Open Findings Count", ""),
        ("Dashboard Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "" or (isinstance(value, str) and not value.startswith("=")):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "PREPARED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Role/Title", "Date"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (Information Security Manager)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Date", "Signature"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO / Executive Management)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Date", "Approval Decision", "Signature"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add(ws[f"B{row-2}"])

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A3"


def main() -> int:
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/8] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/8] Creating Executive_Summary sheet...")
        create_executive_summary_sheet(wb["Executive_Summary"], styles)

        logger.info("[3/8] Creating Review_Status sheet...")
        create_review_status_sheet(wb["Review_Status"], styles)

        logger.info("[4/8] Creating Compliance_Status sheet...")
        create_compliance_status_sheet(wb["Compliance_Status"], styles)

        logger.info("[5/8] Creating Findings_Overview sheet...")
        create_findings_overview_sheet(wb["Findings_Overview"], styles)

        logger.info("[6/8] Creating KPI_Scorecard sheet...")
        create_kpi_scorecard_sheet(wb["KPI_Scorecard"], styles)

        logger.info("[7/8] Creating Trend_Analysis sheet...")
        create_trend_analysis_sheet(wb["Trend_Analysis"], styles)

        logger.info("[8/8] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)
        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.35-36.S4 specification
# =============================================================================
