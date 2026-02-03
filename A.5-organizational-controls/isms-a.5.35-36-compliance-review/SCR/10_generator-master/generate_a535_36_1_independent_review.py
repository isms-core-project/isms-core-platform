#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.35-36.1 - Independent Review Planning & Tracking Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.35: Independent Review of Information Security

Assessment Domain 1 of 4: Independent Review Planning & Tracking

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel workbook for planning, scheduling,
and tracking independent reviews of the information security management system.

**Purpose:**
Enables systematic planning and tracking of independent ISMS reviews as required
by ISO 27001:2022 Control A.5.35, ensuring the organisation's approach to
managing information security is reviewed at planned intervals.

**Generated Workbook Structure:**
1. Instructions - Review planning guidance
2. Review_Schedule - Multi-year review schedule
3. Reviewer_Registry - Qualified reviewer tracking
4. Review_Scope - Scope definition per review
5. Review_Execution - In-progress review tracking
6. Findings_Summary - Review findings overview
7. Evidence_Register - Audit evidence tracking
8. Approval_SignOff - Stakeholder approval

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.35-36.1"
WORKBOOK_NAME = "Independent Review Planning and Tracking"
CONTROL_ID = "A.5.35"
CONTROL_NAME = "Independent Review of Information Security"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
    }
    return styles


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Review_Schedule",
        "Reviewer_Registry",
        "Review_Scope",
        "Review_Execution",
        "Findings_Summary",
        "Evidence_Register",
        "Approval_SignOff",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Independent Review Planning & Tracking"),
        ("Control Reference", CONTROL_ID),
        ("Version", "1.0"),
        ("Planning Period", ""),
        ("Prepared By", ""),
        ("Organisation", ""),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "CONTROL REQUIREMENT"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    ws[f"A{row}"] = (
        "ISO 27001:2022 A.5.35 requires that the organisation's approach to managing "
        "information security and its implementation including people, processes and "
        "technologies should be reviewed independently at planned intervals, or when "
        "significant changes occur."
    )
    ws.merge_cells(f"A{row}:G{row}")

    row += 2
    ws[f"A{row}"] = "REVIEW TRIGGERS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    triggers = [
        ("Planned Interval", "Annual comprehensive review (minimum)"),
        ("Significant Change", "Major organisational restructuring"),
        ("Incident-Driven", "After significant security incident"),
        ("Regulatory Change", "New compliance requirements"),
        ("Technology Change", "Major system implementations"),
        ("Certification", "Pre-certification or surveillance audit"),
    ]

    row += 1
    for trigger, desc in triggers:
        ws[f"A{row}"] = trigger
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = desc
        row += 1

    row += 1
    ws[f"A{row}"] = "INDEPENDENCE REQUIREMENTS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    ws[f"A{row}"] = (
        "Reviewers must be independent of the area being reviewed. Independence can be "
        "achieved through: (a) External auditors/consultants, (b) Internal audit function, "
        "(c) Cross-departmental review (e.g., HR reviewing IT, IT reviewing Finance)."
    )
    ws.merge_cells(f"A{row}:G{row}")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 55
    ws.freeze_panes = "A4"


# =============================================================================
# REVIEW SCHEDULE SHEET
# =============================================================================
def create_review_schedule_sheet(ws, styles):
    """Create the Review_Schedule sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "INDEPENDENT REVIEW SCHEDULE"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Review_ID", 14),
        ("Review_Type", 18),
        ("Review_Scope", 35),
        ("Planned_Start", 14),
        ("Planned_End", 14),
        ("Actual_Start", 14),
        ("Actual_End", 14),
        ("Lead_Reviewer", 22),
        ("Review_Status", 16),
        ("Findings_Count", 14),
        ("Report_Date", 14),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Annual Comprehensive,Surveillance,Ad-Hoc,Pre-Certification,Post-Incident,Scope Change"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_status = DataValidation(
        type="list",
        formula1='"Planned,In Progress,Completed,Postponed,Cancelled"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    # Pre-populate with sample schedule
    reviews = [
        ("REV-2025-001", "Annual Comprehensive", "Full ISMS scope"),
        ("REV-2025-002", "Surveillance", "Selected controls sample"),
        ("REV-2026-001", "Annual Comprehensive", "Full ISMS scope"),
        ("REV-2026-002", "Surveillance", "Selected controls sample"),
    ]

    for row_idx, (rev_id, rev_type, scope) in enumerate(reviews, start=4):
        ws.cell(row=row_idx, column=1, value=rev_id)
        ws.cell(row=row_idx, column=2, value=rev_type)
        ws.cell(row=row_idx, column=3, value=scope)

        for c in range(4, 13):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=row_idx, column=2))
        dv_status.add(ws.cell(row=row_idx, column=9))

    # Additional blank rows
    for r in range(len(reviews) + 4, len(reviews) + 24):
        ws.cell(row=r, column=1, value=f"REV-XXXX-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=2))
        dv_status.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "C4"


# =============================================================================
# REVIEWER REGISTRY SHEET
# =============================================================================
def create_reviewer_registry_sheet(ws, styles):
    """Create the Reviewer_Registry sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "QUALIFIED REVIEWER REGISTRY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Reviewer_ID", 14),
        ("Reviewer_Name", 25),
        ("Organisation", 25),
        ("Reviewer_Type", 16),
        ("Qualifications", 35),
        ("Certification_Expiry", 16),
        ("Areas_of_Expertise", 35),
        ("Independence_Status", 18),
        ("Last_Review_Date", 16),
        ("Approval_Status", 16),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"External Auditor,External Consultant,Internal Audit,Cross-Departmental"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_independence = DataValidation(
        type="list",
        formula1='"Fully Independent,Conditionally Independent,Not Independent"',
        allow_blank=False
    )
    ws.add_data_validation(dv_independence)

    dv_approval = DataValidation(
        type="list",
        formula1='"Approved,Pending Approval,Expired,Revoked"',
        allow_blank=False
    )
    ws.add_data_validation(dv_approval)

    for r in range(4, 24):
        ws.cell(row=r, column=1, value=f"RVR-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 12):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=4))
        dv_independence.add(ws.cell(row=r, column=8))
        dv_approval.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "C4"


# =============================================================================
# REVIEW SCOPE SHEET
# =============================================================================
def create_review_scope_sheet(ws, styles):
    """Create the Review_Scope sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "REVIEW SCOPE DEFINITION"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Review_Ref", 14),
        ("Scope_Area", 30),
        ("ISO_Controls", 25),
        ("Departments_Included", 25),
        ("Systems_Included", 30),
        ("Exclusions", 25),
        ("Sampling_Approach", 22),
        ("Sample_Size", 14),
        ("Scope_Approved_By", 20),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate common scope areas
    scope_areas = [
        ("ISMS Governance", "A.5.1-A.5.4", "Executive, IS"),
        ("Risk Management", "A.5.5-A.5.6", "IS, Risk"),
        ("Access Control", "A.5.15-A.5.18, A.8.2-A.8.5", "IT, IS"),
        ("Operations Security", "A.8.9-A.8.22", "IT Ops"),
        ("Incident Management", "A.5.24-A.5.28", "IS, IT"),
        ("Business Continuity", "A.5.29-A.5.30, A.8.13-A.8.14", "BCM, IT"),
        ("Compliance", "A.5.31-A.5.36", "Legal, IS"),
        ("Human Resources", "A.6.1-A.6.8", "HR, IS"),
        ("Physical Security", "A.7.1-A.7.14", "Facilities"),
    ]

    dv_sampling = DataValidation(
        type="list",
        formula1='"Statistical,Judgmental,100% Coverage,Risk-Based"',
        allow_blank=False
    )
    ws.add_data_validation(dv_sampling)

    for row_idx, (area, controls, depts) in enumerate(scope_areas, start=4):
        ws.cell(row=row_idx, column=2, value=area)
        ws.cell(row=row_idx, column=3, value=controls)
        ws.cell(row=row_idx, column=4, value=depts)

        for c in [1] + list(range(5, 11)):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_sampling.add(ws.cell(row=row_idx, column=7))

    # Additional blank rows
    for r in range(len(scope_areas) + 4, len(scope_areas) + 14):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_sampling.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "C4"


# =============================================================================
# REVIEW EXECUTION SHEET
# =============================================================================
def create_review_execution_sheet(ws, styles):
    """Create the Review_Execution sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "REVIEW EXECUTION TRACKING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Activity_ID", 14),
        ("Review_Ref", 14),
        ("Activity_Description", 40),
        ("Assigned_To", 22),
        ("Planned_Date", 14),
        ("Actual_Date", 14),
        ("Status", 16),
        ("Evidence_Collected", 16),
        ("Interviews_Conducted", 18),
        ("Issues_Found", 14),
        ("Follow_Up_Required", 16),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Completed,Blocked"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No,Partial"',
        allow_blank=False
    )
    ws.add_data_validation(dv_yesno)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"ACT-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_status.add(ws.cell(row=r, column=7))
        dv_yesno.add(ws.cell(row=r, column=8))
        dv_yesno.add(ws.cell(row=r, column=11))

    ws.freeze_panes = "D4"


# =============================================================================
# FINDINGS SUMMARY SHEET
# =============================================================================
def create_findings_summary_sheet(ws, styles):
    """Create the Findings_Summary sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "REVIEW FINDINGS SUMMARY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Finding_ID", 14),
        ("Review_Ref", 14),
        ("Finding_Type", 18),
        ("Finding_Description", 50),
        ("Control_Reference", 18),
        ("Severity", 14),
        ("Root_Cause", 30),
        ("Recommendation", 40),
        ("Management_Response", 35),
        ("Remediation_Owner", 22),
        ("Target_Date", 14),
        ("Status", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Non-Conformity,Observation,Opportunity for Improvement,Positive Finding"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_severity = DataValidation(
        type="list",
        formula1='"Major,Minor,Observation"',
        allow_blank=False
    )
    ws.add_data_validation(dv_severity)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Closed,Accepted"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"FND-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=3))
        dv_severity.add(ws.cell(row=r, column=6))
        dv_status.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "D4"


# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================
def create_evidence_register_sheet(ws, styles):
    """Create the Evidence_Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 45),
        ("Related_Review", 18),
        ("Collection_Date", 16),
        ("Location", 40),
        ("Collected_By", 25),
        ("Retention_Until", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Review Plan,Scope Document,Interview Notes,Sample Evidence,Audit Report,Management Response,Closure Evidence"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "C4"


# =============================================================================
# APPROVAL SIGN-OFF SHEET
# =============================================================================
def create_approval_signoff_sheet(ws, styles):
    """Create the Approval_SignOff sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "REVIEW PLANNING APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Review Planning Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Planning Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Planning Period", ""),
        ("Total Reviews Scheduled", "=COUNTA(Review_Schedule!A4:A27)-COUNTBLANK(Review_Schedule!B4:B27)"),
        ("Approved Reviewers", '=COUNTIF(Reviewer_Registry!J4:J23,"Approved")'),
        ("Planning Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "" or (isinstance(value, str) and not value.startswith("=")):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "PREPARED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Role/Title", "Date"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Date", "Approval Decision", "Signature"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add(ws[f"B{row-2}"])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A3"


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/8] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/8] Creating Review_Schedule sheet...")
        create_review_schedule_sheet(wb["Review_Schedule"], styles)

        logger.info("[3/8] Creating Reviewer_Registry sheet...")
        create_reviewer_registry_sheet(wb["Reviewer_Registry"], styles)

        logger.info("[4/8] Creating Review_Scope sheet...")
        create_review_scope_sheet(wb["Review_Scope"], styles)

        logger.info("[5/8] Creating Review_Execution sheet...")
        create_review_execution_sheet(wb["Review_Execution"], styles)

        logger.info("[6/8] Creating Findings_Summary sheet...")
        create_findings_summary_sheet(wb["Findings_Summary"], styles)

        logger.info("[7/8] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[8/8] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)
        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.35-36.1 specification
# =============================================================================
