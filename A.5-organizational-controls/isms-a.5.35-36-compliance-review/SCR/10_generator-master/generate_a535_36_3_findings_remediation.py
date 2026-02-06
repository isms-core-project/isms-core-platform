#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.35-36.S3 - Findings & Remediation Management Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.5.35-36: Compliance & Review

Assessment Domain 3 of 4: Findings & Remediation Management

This script generates a workbook for consolidated findings tracking and
remediation management from both independent reviews (A.5.35) and
compliance assessments (A.5.36).

**Generated Workbook Structure:**
1. Instructions - Findings management guidance
2. Findings_Register - Consolidated findings from all reviews
3. Remediation_Actions - Action item tracking
4. Root_Cause_Analysis - RCA for significant findings
5. Verification_Log - Closure verification tracking
6. Trend_Analysis - Findings trends over time
7. Evidence_Register - Audit evidence
8. Approval_SignOff - Stakeholder approval

================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

DOCUMENT_ID = "ISMS-IMP-A.5.35-36.S3"
WORKBOOK_NAME = "Findings and Remediation Management"
CONTROL_ID = "A.5.35-36"
CONTROL_NAME = "Compliance & Review"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


def setup_styles():
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
    }


def create_workbook() -> Workbook:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    for name in ["Instructions", "Findings_Register", "Remediation_Actions",
                 "Root_Cause_Analysis", "Verification_Log", "Trend_Analysis",
                 "Evidence_Register", "Approval_SignOff"]:
        wb.create_sheet(title=name)
    return wb


def create_instructions_sheet(ws, styles):
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Findings & Remediation Management"),
        ("Control Reference", CONTROL_ID),
        ("Version", "1.0"),
        ("Tracking Period", ""),
        ("Managed By", ""),
        ("Organisation", ""),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "PURPOSE"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    ws[f"A{row}"] = (
        "This workbook consolidates findings from independent reviews (A.5.35) and "
        "compliance assessments (A.5.36), tracking remediation through to verified closure."
    )
    ws.merge_cells(f"A{row}:G{row}")

    row += 2
    ws[f"A{row}"] = "FINDING SEVERITY DEFINITIONS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    severity_definitions = [
        ("Major NC", "Significant gap; absence or total failure of control; immediate action required"),
        ("Minor NC", "Partial gap; control exists but not fully effective; corrective action needed"),
        ("Observation", "Opportunity for improvement; not a compliance failure"),
        ("Positive", "Noteworthy good practice identified"),
    ]

    row += 1
    for severity, desc in severity_definitions:
        ws[f"A{row}"] = severity
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = desc
        row += 1

    row += 1
    ws[f"A{row}"] = "REMEDIATION TIMELINE REQUIREMENTS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    timelines = [
        ("Major NC", "30 days maximum (immediate containment within 48 hours)"),
        ("Minor NC", "90 days maximum"),
        ("Observation", "Next assessment cycle"),
    ]

    row += 1
    for finding_type, timeline in timelines:
        ws[f"A{row}"] = finding_type
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = timeline
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
    ws.freeze_panes = "A4"


def create_findings_register_sheet(ws, styles):
    ws.merge_cells("A1:N1")
    ws["A1"] = "CONSOLIDATED FINDINGS REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Finding_ID", 14),
        ("Source", 18),
        ("Source_Ref", 16),
        ("Finding_Date", 14),
        ("Finding_Type", 16),
        ("Severity", 14),
        ("Control_Reference", 18),
        ("Finding_Description", 50),
        ("Root_Cause", 35),
        ("Recommendation", 40),
        ("Owner", 22),
        ("Target_Date", 14),
        ("Status", 16),
        ("Days_Open", 12),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_source = DataValidation(
        type="list",
        formula1='"Independent Review,Compliance Assessment,Internal Audit,Certification Audit,Surveillance Audit,Management Review"',
        allow_blank=False
    )
    ws.add_data_validation(dv_source)

    dv_type = DataValidation(
        type="list",
        formula1='"Non-Conformity,Observation,Opportunity for Improvement,Positive Finding"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_severity = DataValidation(
        type="list",
        formula1='"Major,Minor,Observation,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_severity)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Pending Verification,Closed,Risk Accepted"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"FND-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 15):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Days open formula
        ws.cell(row=r, column=14).value = f'=IF(AND(D{r}<>"",M{r}<>"Closed"),TODAY()-D{r},"")'
        ws.cell(row=r, column=14).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        dv_source.add(ws.cell(row=r, column=2))
        dv_type.add(ws.cell(row=r, column=5))
        dv_severity.add(ws.cell(row=r, column=6))
        dv_status.add(ws.cell(row=r, column=13))

    ws.freeze_panes = "D4"


def create_remediation_actions_sheet(ws, styles):
    ws.merge_cells("A1:L1")
    ws["A1"] = "REMEDIATION ACTION TRACKER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Action_ID", 14),
        ("Finding_Ref", 14),
        ("Action_Description", 50),
        ("Action_Type", 18),
        ("Owner", 22),
        ("Start_Date", 14),
        ("Target_Date", 14),
        ("Actual_Date", 14),
        ("Status", 16),
        ("% Complete", 12),
        ("Verification_Method", 25),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Corrective Action,Preventive Action,Immediate Containment,Process Improvement"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Completed,On Hold,Cancelled"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 204):
        ws.cell(row=r, column=1, value=f"ACT-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=4))
        dv_status.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "D4"


def create_root_cause_analysis_sheet(ws, styles):
    ws.merge_cells("A1:K1")
    ws["A1"] = "ROOT CAUSE ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("RCA_ID", 12),
        ("Finding_Ref", 14),
        ("Problem_Statement", 40),
        ("Analysis_Method", 18),
        ("Root_Cause_Category", 22),
        ("Root_Cause_Description", 45),
        ("Contributing_Factors", 35),
        ("Systemic_Issue", 14),
        ("Preventive_Actions", 40),
        ("Analyst", 22),
        ("Analysis_Date", 14),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_method = DataValidation(
        type="list",
        formula1='"5 Whys,Fishbone Diagram,Fault Tree,Pareto Analysis,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_method)

    dv_category = DataValidation(
        type="list",
        formula1='"People,Process,Technology,Governance,Training,Resources,External"',
        allow_blank=False
    )
    ws.add_data_validation(dv_category)

    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_yesno)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"RCA-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 12):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_method.add(ws.cell(row=r, column=4))
        dv_category.add(ws.cell(row=r, column=5))
        dv_yesno.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "C4"


def create_verification_log_sheet(ws, styles):
    ws.merge_cells("A1:K1")
    ws["A1"] = "VERIFICATION LOG"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Verification_ID", 14),
        ("Finding_Ref", 14),
        ("Action_Ref", 14),
        ("Verification_Date", 16),
        ("Verifier", 22),
        ("Verification_Method", 25),
        ("Evidence_Reviewed", 35),
        ("Verification_Result", 18),
        ("Closure_Recommendation", 22),
        ("Follow_Up_Required", 16),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_method = DataValidation(
        type="list",
        formula1='"Document Review,Testing,Interview,Observation,System Check,Sample Verification"',
        allow_blank=False
    )
    ws.add_data_validation(dv_method)

    dv_result = DataValidation(
        type="list",
        formula1='"Effective,Partially Effective,Not Effective"',
        allow_blank=False
    )
    ws.add_data_validation(dv_result)

    dv_closure = DataValidation(
        type="list",
        formula1='"Close,Keep Open,Re-work Required"',
        allow_blank=False
    )
    ws.add_data_validation(dv_closure)

    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_yesno)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"VER-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 12):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_method.add(ws.cell(row=r, column=6))
        dv_result.add(ws.cell(row=r, column=8))
        dv_closure.add(ws.cell(row=r, column=9))
        dv_yesno.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "D4"


def create_trend_analysis_sheet(ws, styles):
    ws.merge_cells("A1:J1")
    ws["A1"] = "FINDINGS TREND ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Period", 16),
        ("Major_NCs", 12),
        ("Minor_NCs", 12),
        ("Observations", 14),
        ("Total_Findings", 14),
        ("Closed_This_Period", 16),
        ("Open_at_Period_End", 18),
        ("Avg_Days_to_Close", 16),
        ("Repeat_Findings", 14),
        ("Notes", 35),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate periods
    periods = ["Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025",
               "Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"]

    for row_idx, period in enumerate(periods, start=4):
        ws.cell(row=row_idx, column=1, value=period)
        for c in range(2, 11):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Total formula
        ws.cell(row=row_idx, column=5).value = f"=SUM(B{row_idx}:D{row_idx})"
        ws.cell(row=row_idx, column=5).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.freeze_panes = "B4"


def create_evidence_register_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 45),
        ("Related_Finding", 18),
        ("Collection_Date", 16),
        ("Location", 40),
        ("Collected_By", 25),
        ("Retention_Until", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Finding Report,Remediation Plan,Closure Evidence,RCA Document,Verification Record,Management Approval"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "C4"


def create_approval_signoff_sheet(ws, styles):
    ws.merge_cells("A1:E1")
    ws["A1"] = "FINDINGS MANAGEMENT APPROVAL"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Findings Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Management Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Reporting Period", ""),
        ("Total Findings", "=COUNTA(Findings_Register!A4:A103)-COUNTBLANK(Findings_Register!B4:B103)"),
        ("Open Major NCs", '=COUNTIFS(Findings_Register!F4:F103,"Major",Findings_Register!M4:M103,"Open")'),
        ("Open Minor NCs", '=COUNTIFS(Findings_Register!F4:F103,"Minor",Findings_Register!M4:M103,"Open")'),
        ("Findings Closed This Period", '=COUNTIF(Findings_Register!M4:M103,"Closed")'),
        ("Actions In Progress", '=COUNTIF(Remediation_Actions!I4:I203,"In Progress")'),
        ("Management Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "" or (isinstance(value, str) and not value.startswith("=")):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "PREPARED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Role/Title", "Date"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Date", "Approval Decision", "Signature"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add(ws[f"B{row-2}"])

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A3"


def main() -> int:
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/8] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/8] Creating Findings_Register sheet...")
        create_findings_register_sheet(wb["Findings_Register"], styles)

        logger.info("[3/8] Creating Remediation_Actions sheet...")
        create_remediation_actions_sheet(wb["Remediation_Actions"], styles)

        logger.info("[4/8] Creating Root_Cause_Analysis sheet...")
        create_root_cause_analysis_sheet(wb["Root_Cause_Analysis"], styles)

        logger.info("[5/8] Creating Verification_Log sheet...")
        create_verification_log_sheet(wb["Verification_Log"], styles)

        logger.info("[6/8] Creating Trend_Analysis sheet...")
        create_trend_analysis_sheet(wb["Trend_Analysis"], styles)

        logger.info("[7/8] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[8/8] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)
        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.35-36.S3 specification
# =============================================================================
