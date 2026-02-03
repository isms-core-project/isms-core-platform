#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.35-36.2 - Compliance Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.36: Compliance with Policies, Rules and Standards

Assessment Domain 2 of 4: Compliance Assessment

This script generates a workbook for operational compliance verification,
enabling managers to assess their areas against information security policies.

**Generated Workbook Structure:**
1. Instructions - Compliance assessment guidance
2. Policy_Compliance - Policy-by-policy compliance checklist
3. Control_Compliance - Control-by-control compliance status
4. Department_Assessment - Department self-assessment
5. NonCompliance_Register - Non-compliance tracking
6. Evidence_Register - Audit evidence tracking
7. Approval_SignOff - Stakeholder approval

================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

DOCUMENT_ID = "ISMS-IMP-A.5.35-36.2"
WORKBOOK_NAME = "Compliance Assessment"
CONTROL_ID = "A.5.36"
CONTROL_NAME = "Compliance with Policies, Rules and Standards"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


def setup_styles():
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }


def create_workbook() -> Workbook:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    for name in ["Instructions", "Policy_Compliance", "Control_Compliance",
                 "Department_Assessment", "NonCompliance_Register",
                 "Evidence_Register", "Approval_SignOff"]:
        wb.create_sheet(title=name)
    return wb


def create_instructions_sheet(ws, styles):
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Operational Compliance Assessment"),
        ("Control Reference", CONTROL_ID),
        ("Version", "1.0"),
        ("Assessment Period", ""),
        ("Assessed By", ""),
        ("Organisation", ""),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "CONTROL REQUIREMENT"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    ws[f"A{row}"] = (
        "ISO 27001:2022 A.5.36 requires that compliance with the organisation's "
        "information security policy, topic-specific policies, rules and standards "
        "should be regularly reviewed."
    )
    ws.merge_cells(f"A{row}:G{row}")

    row += 2
    ws[f"A{row}"] = "ASSESSMENT APPROACH"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    approach = [
        ("Manager Self-Assessment", "Line managers verify compliance in their areas"),
        ("Evidence-Based", "Claims supported by documented evidence"),
        ("Regular Cadence", "Quarterly operational reviews recommended"),
        ("Non-Compliance Tracking", "All non-compliance logged and remediated"),
        ("Escalation Process", "Persistent issues escalated to management"),
    ]

    row += 1
    for item, desc in approach:
        ws[f"A{row}"] = item
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = desc
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    ws.freeze_panes = "A4"


def create_policy_compliance_sheet(ws, styles):
    ws.merge_cells("A1:K1")
    ws["A1"] = "POLICY COMPLIANCE ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Policy_ID", 16),
        ("Policy_Name", 40),
        ("Policy_Version", 14),
        ("Compliance_Status", 18),
        ("Last_Reviewed", 14),
        ("Assessed_By", 22),
        ("Assessment_Date", 14),
        ("Evidence_Ref", 18),
        ("NonCompliance_Issues", 30),
        ("Remediation_Status", 18),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate with common ISMS policies
    policies = [
        ("ISMS-POL-00", "Regulatory Applicability Framework"),
        ("ISMS-POL-A.5.1-4", "Information Security Governance"),
        ("ISMS-POL-A.5.10-11", "Asset Usage Lifecycle"),
        ("ISMS-POL-A.5.15-18", "Identity & Access Management"),
        ("ISMS-POL-A.5.24-28", "Incident Management"),
        ("ISMS-POL-A.6.1-5", "Human Resources Security"),
        ("ISMS-POL-A.7.1-14", "Physical Security"),
        ("ISMS-POL-A.8.1-7", "Endpoint Security"),
        ("ISMS-POL-A.8.9", "Configuration Management"),
        ("ISMS-POL-A.8.15-16", "Logging & Monitoring"),
        ("ISMS-POL-A.8.20-22", "Network Security"),
        ("ISMS-POL-A.8.24", "Cryptography"),
        ("ISMS-POL-A.8.25-31", "Secure Development"),
    ]

    dv_status = DataValidation(
        type="list",
        formula1='"Compliant,Partial,Non-Compliant,Not Assessed,Not Applicable"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_remediation = DataValidation(
        type="list",
        formula1='"N/A,Open,In Progress,Closed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_remediation)

    for row_idx, (pol_id, pol_name) in enumerate(policies, start=4):
        ws.cell(row=row_idx, column=1, value=pol_id)
        ws.cell(row=row_idx, column=2, value=pol_name)
        for c in range(3, 12):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_status.add(ws.cell(row=row_idx, column=4))
        dv_remediation.add(ws.cell(row=row_idx, column=10))

    # Additional blank rows
    for r in range(len(policies) + 4, len(policies) + 24):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_status.add(ws.cell(row=r, column=4))
        dv_remediation.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "C4"


def create_control_compliance_sheet(ws, styles):
    ws.merge_cells("A1:J1")
    ws["A1"] = "CONTROL COMPLIANCE STATUS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Control_ID", 14),
        ("Control_Name", 45),
        ("Control_Category", 22),
        ("Compliance_Status", 18),
        ("Implementation_%", 14),
        ("Last_Assessed", 14),
        ("Assessed_By", 22),
        ("Evidence_Ref", 18),
        ("Gaps_Identified", 30),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ISO 27001:2022 Annex A control categories
    control_categories = [
        ("A.5", "Organisational Controls", "Organisational"),
        ("A.6", "People Controls", "People"),
        ("A.7", "Physical Controls", "Physical"),
        ("A.8", "Technological Controls", "Technological"),
    ]

    dv_status = DataValidation(
        type="list",
        formula1='"Fully Implemented,Partially Implemented,Not Implemented,Not Applicable"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for row_idx, (ctrl_id, ctrl_name, category) in enumerate(control_categories, start=4):
        ws.cell(row=row_idx, column=1, value=ctrl_id)
        ws.cell(row=row_idx, column=2, value=ctrl_name)
        ws.cell(row=row_idx, column=3, value=category)
        for c in range(4, 11):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_status.add(ws.cell(row=row_idx, column=4))

    # Additional blank rows for detailed controls
    for r in range(len(control_categories) + 4, 100):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_status.add(ws.cell(row=r, column=4))

    ws.freeze_panes = "C4"


def create_department_assessment_sheet(ws, styles):
    ws.merge_cells("A1:L1")
    ws["A1"] = "DEPARTMENT SELF-ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Department", 25),
        ("Manager", 22),
        ("Assessment_Period", 16),
        ("AUP_Compliance", 16),
        ("Access_Control_Compliance", 20),
        ("Incident_Reporting_Compliance", 22),
        ("Training_Compliance", 18),
        ("Asset_Management_Compliance", 22),
        ("Overall_Rating", 16),
        ("Issues_Identified", 30),
        ("Improvement_Actions", 30),
        ("Sign_Off_Date", 14),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    departments = [
        "Executive Office", "Finance", "Human Resources", "IT Operations",
        "Software Development", "Sales", "Marketing", "Customer Support",
        "Legal", "Facilities", "Research & Development",
    ]

    dv_compliance = DataValidation(
        type="list",
        formula1='"Compliant,Partial,Non-Compliant,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_compliance)

    dv_rating = DataValidation(
        type="list",
        formula1='"Green,Amber,Red"',
        allow_blank=False
    )
    ws.add_data_validation(dv_rating)

    for row_idx, dept in enumerate(departments, start=4):
        ws.cell(row=row_idx, column=1, value=dept)
        for c in range(2, 13):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        for c in [4, 5, 6, 7, 8]:
            dv_compliance.add(ws.cell(row=row_idx, column=c))
        dv_rating.add(ws.cell(row=row_idx, column=9))

    # Additional blank rows
    for r in range(len(departments) + 4, len(departments) + 14):
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        for c in [4, 5, 6, 7, 8]:
            dv_compliance.add(ws.cell(row=r, column=c))
        dv_rating.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "B4"


def create_noncompliance_register_sheet(ws, styles):
    ws.merge_cells("A1:L1")
    ws["A1"] = "NON-COMPLIANCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("NC_ID", 12),
        ("Identified_Date", 14),
        ("Policy_Control_Ref", 20),
        ("Department", 20),
        ("NC_Description", 45),
        ("Root_Cause", 30),
        ("Severity", 14),
        ("Remediation_Action", 35),
        ("Owner", 22),
        ("Target_Date", 14),
        ("Status", 16),
        ("Closure_Date", 14),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_severity = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_severity)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Pending Verification,Closed,Risk Accepted"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"NC-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_severity.add(ws.cell(row=r, column=7))
        dv_status.add(ws.cell(row=r, column=11))

    ws.freeze_panes = "D4"


def create_evidence_register_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 45),
        ("Related_Assessment", 20),
        ("Collection_Date", 16),
        ("Location", 40),
        ("Collected_By", 25),
        ("Valid_Until", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Policy Document,Configuration Export,Training Record,Attestation,Screenshot,Log Extract,Report"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "C4"


def create_approval_signoff_sheet(ws, styles):
    ws.merge_cells("A1:E1")
    ws["A1"] = "COMPLIANCE ASSESSMENT APPROVAL"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Policies Assessed", "=COUNTA(Policy_Compliance!A4:A36)-COUNTBLANK(Policy_Compliance!B4:B36)"),
        ("Policies Compliant", '=COUNTIF(Policy_Compliance!D4:D36,"Compliant")'),
        ("Open Non-Conformances", '=COUNTIF(NonCompliance_Register!K4:K103,"Open")'),
        ("Assessment Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "" or (isinstance(value, str) and not value.startswith("=")):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Role/Title", "Date"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (Information Security Manager)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Date", "Signature"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Date", "Approval Decision", "Signature"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add(ws[f"B{row-2}"])

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A3"


def main() -> int:
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/7] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/7] Creating Policy_Compliance sheet...")
        create_policy_compliance_sheet(wb["Policy_Compliance"], styles)

        logger.info("[3/7] Creating Control_Compliance sheet...")
        create_control_compliance_sheet(wb["Control_Compliance"], styles)

        logger.info("[4/7] Creating Department_Assessment sheet...")
        create_department_assessment_sheet(wb["Department_Assessment"], styles)

        logger.info("[5/7] Creating NonCompliance_Register sheet...")
        create_noncompliance_register_sheet(wb["NonCompliance_Register"], styles)

        logger.info("[6/7] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[7/7] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)
        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.35-36.2 specification
# =============================================================================
