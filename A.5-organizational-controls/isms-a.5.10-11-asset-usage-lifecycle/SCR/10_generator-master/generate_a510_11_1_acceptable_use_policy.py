#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.10-11.1 - Acceptable Use Policy Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.10: Acceptable Use of Information and Other
Associated Assets

Assessment Domain 1 of 4: Acceptable Use Policy Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific acceptable use requirements, asset categories,
and policy structure.

Key customization areas:
1. Asset categories (adapt to your organizational structure)
2. Policy requirements (align with your governance framework)
3. Communication channels (based on your awareness program)
4. Role definitions (match your access control matrix)
5. Evidence requirements (per your audit standards)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-IMP-A.5.10-11.1 specification

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
acceptable use policies and their coverage across information assets.

**Purpose:**
Enables systematic assessment of Acceptable Use Policy (AUP) completeness,
coverage, communication, and compliance against ISO 27001:2022 Control A.5.10
requirements.

**Assessment Scope:**
- AUP policy content completeness
- Asset category coverage
- User awareness and acknowledgment tracking
- Policy communication effectiveness
- Enforcement and monitoring mechanisms
- Exception handling processes

**Generated Workbook Structure:**
1. Instructions - Assessment guidance and methodology
2. Policy_Assessment - AUP completeness evaluation
3. Asset_Coverage - Coverage by asset category
4. Awareness_Tracking - User acknowledgment tracking
5. Communication_Matrix - Policy dissemination assessment
6. Evidence_Register - Audit evidence tracking
7. Approval_SignOff - Stakeholder review and approval

================================================================================
"""

# =============================================================================
# IMPORTS - Standard Library
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# IMPORTS - Third Party
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.10-11.1"
WORKBOOK_NAME = "Acceptable Use Policy Assessment"
CONTROL_ID = "A.5.10"
CONTROL_NAME = "Acceptable Use of Information and Other Associated Assets"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure matches specification
    sheets = [
        "Instructions",
        "Policy_Assessment",
        "Asset_Coverage",
        "Awareness_Tracking",
        "Communication_Matrix",
        "Evidence_Register",
        "Approval_SignOff",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet with assessment guidance."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Acceptable Use Policy Completeness"),
        ("Control Reference", CONTROL_ID),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Annual"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "CONTROL REQUIREMENT"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    ws[f"A{row}"] = (
        "ISO 27001:2022 A.5.10 requires that rules for the acceptable use of "
        "information and other associated assets should be identified, documented "
        "and implemented."
    )
    ws.merge_cells(f"A{row}:G{row}")

    row += 2
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Complete Policy_Assessment to evaluate AUP content completeness",
        "2. Review Asset_Coverage to verify all asset categories are addressed",
        "3. Track user acknowledgments in Awareness_Tracking",
        "4. Assess communication effectiveness in Communication_Matrix",
        "5. Link supporting evidence in Evidence_Register",
        "6. Obtain approvals in Approval_SignOff sheet",
    ]

    row += 1
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    row += 1
    ws[f"A{row}"] = "AUP REQUIRED CONTENT AREAS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    content_areas = [
        ("Permitted Use", "Clear definition of acceptable activities"),
        ("Prohibited Use", "Explicit list of forbidden activities"),
        ("Personal Use", "Guidelines for personal use of company assets"),
        ("Monitoring Notice", "Disclosure of monitoring activities"),
        ("Consequences", "Disciplinary actions for violations"),
        ("Exceptions", "Process for requesting policy exceptions"),
        ("Intellectual Property", "IP ownership and handling rules"),
        ("Data Classification", "Handling requirements by classification"),
    ]

    row += 1
    for area, desc in content_areas:
        ws[f"A{row}"] = area
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = desc
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    ws.freeze_panes = "A4"


# =============================================================================
# POLICY ASSESSMENT SHEET
# =============================================================================
def create_policy_assessment_sheet(ws, styles):
    """Create the Policy_Assessment sheet - AUP completeness evaluation."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "ACCEPTABLE USE POLICY ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Column definitions
    columns = [
        ("Requirement_ID", 15),
        ("Policy_Requirement", 45),
        ("Category", 20),
        ("Addressed", 14),
        ("Policy_Section", 20),
        ("Clarity_Rating", 16),
        ("Last_Updated", 16),
        ("Owner", 22),
        ("Gap_Status", 16),
        ("Remediation_Notes", 35),
        ("Evidence_Ref", 18),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate with ISO 27001:2022 A.5.10 requirements
    # CUSTOMIZE: Adapt these requirements to your organization
    requirements = [
        ("REQ-001", "Policy scope defines applicable assets", "Scope", "Information systems, endpoints, data"),
        ("REQ-002", "Policy defines acceptable business use", "Permitted Use", "Business email, approved applications"),
        ("REQ-003", "Policy explicitly prohibits unauthorized activities", "Prohibited Use", "Unauthorized software, data exfiltration"),
        ("REQ-004", "Personal use guidelines are documented", "Personal Use", "Limited personal use provisions"),
        ("REQ-005", "Monitoring and logging disclosure included", "Monitoring", "User awareness of monitoring"),
        ("REQ-006", "Consequences of violation are stated", "Enforcement", "Disciplinary measures"),
        ("REQ-007", "Exception request process documented", "Exceptions", "How to request exceptions"),
        ("REQ-008", "Intellectual property rules defined", "IP Protection", "Ownership and handling"),
        ("REQ-009", "Data classification handling requirements", "Data Handling", "By classification level"),
        ("REQ-010", "Mobile device acceptable use rules", "Mobile Devices", "BYOD and corporate mobile"),
        ("REQ-011", "Remote working asset use guidelines", "Remote Work", "Home/remote asset handling"),
        ("REQ-012", "Cloud service usage rules", "Cloud Services", "Approved vs unauthorized"),
        ("REQ-013", "Social media usage guidelines", "Social Media", "Business vs personal context"),
        ("REQ-014", "Email and messaging acceptable use", "Communications", "Email, IM, collaboration"),
        ("REQ-015", "Internet usage guidelines", "Internet", "Web browsing, downloads"),
        ("REQ-016", "Password and authentication requirements", "Authentication", "User responsibilities"),
        ("REQ-017", "Physical asset protection responsibilities", "Physical Security", "Laptops, tokens, badges"),
        ("REQ-018", "Incident reporting obligations", "Reporting", "When and how to report"),
        ("REQ-019", "Third-party access provisions", "Third Parties", "Contractor/vendor guidelines"),
        ("REQ-020", "Policy acknowledgment requirement", "Acknowledgment", "Annual sign-off required"),
    ]

    # Data validations
    dv_addressed = DataValidation(
        type="list",
        formula1='"Yes,Partial,No,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_addressed)

    dv_clarity = DataValidation(
        type="list",
        formula1='"Clear,Needs Improvement,Unclear"',
        allow_blank=False
    )
    ws.add_data_validation(dv_clarity)

    dv_gap = DataValidation(
        type="list",
        formula1='"Compliant,Gap Identified,Remediation In Progress"',
        allow_blank=False
    )
    ws.add_data_validation(dv_gap)

    # Populate requirements
    for row_idx, (req_id, req, category, description) in enumerate(requirements, start=4):
        ws.cell(row=row_idx, column=1, value=req_id)
        ws.cell(row=row_idx, column=2, value=req)
        ws.cell(row=row_idx, column=3, value=category)

        for c in range(4, 12):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_addressed.add(ws.cell(row=row_idx, column=4))
        dv_clarity.add(ws.cell(row=row_idx, column=6))
        dv_gap.add(ws.cell(row=row_idx, column=9))

    # Additional blank rows
    for r in range(len(requirements) + 4, len(requirements) + 14):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_addressed.add(ws.cell(row=r, column=4))
        dv_clarity.add(ws.cell(row=r, column=6))
        dv_gap.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "C4"


# =============================================================================
# ASSET COVERAGE SHEET
# =============================================================================
def create_asset_coverage_sheet(ws, styles):
    """Create the Asset_Coverage sheet - coverage by asset category."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "ASSET COVERAGE MATRIX"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Asset_Category", 25),
        ("Asset_Examples", 40),
        ("Covered_In_AUP", 16),
        ("Policy_Section", 20),
        ("Usage_Rules_Defined", 18),
        ("Handling_Rules_Defined", 18),
        ("Gap_Notes", 35),
        ("Remediation_Required", 18),
        ("Evidence_Ref", 18),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate asset categories
    # CUSTOMIZE: Adapt to your organization's asset types
    asset_categories = [
        ("Information Assets", "Documents, databases, intellectual property"),
        ("Software Assets", "Applications, operating systems, development tools"),
        ("Hardware Assets", "Laptops, desktops, servers, mobile devices"),
        ("Network Assets", "Routers, switches, firewalls, VPN"),
        ("Cloud Services", "SaaS, IaaS, PaaS platforms"),
        ("Communication Tools", "Email, Teams, Slack, video conferencing"),
        ("Physical Media", "USB drives, external HDDs, backup tapes"),
        ("Authentication Assets", "Tokens, smart cards, certificates"),
        ("Development Environments", "IDEs, repositories, CI/CD pipelines"),
        ("Monitoring Systems", "SIEM, logging, alerting tools"),
        ("Personal Devices (BYOD)", "Personal phones, tablets, laptops"),
        ("IoT Devices", "Sensors, cameras, smart devices"),
    ]

    # Data validations
    dv_covered = DataValidation(
        type="list",
        formula1='"Yes,Partial,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_covered)

    dv_defined = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_defined)

    # Populate rows
    for row_idx, (category, examples) in enumerate(asset_categories, start=4):
        ws.cell(row=row_idx, column=1, value=category)
        ws.cell(row=row_idx, column=2, value=examples)

        for c in range(3, 10):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_covered.add(ws.cell(row=row_idx, column=3))
        dv_defined.add(ws.cell(row=row_idx, column=5))
        dv_defined.add(ws.cell(row=row_idx, column=6))
        dv_defined.add(ws.cell(row=row_idx, column=8))

    # Additional blank rows
    for r in range(len(asset_categories) + 4, len(asset_categories) + 14):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_covered.add(ws.cell(row=r, column=3))
        dv_defined.add(ws.cell(row=r, column=5))
        dv_defined.add(ws.cell(row=r, column=6))
        dv_defined.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "C4"


# =============================================================================
# AWARENESS TRACKING SHEET
# =============================================================================
def create_awareness_tracking_sheet(ws, styles):
    """Create the Awareness_Tracking sheet - user acknowledgment tracking."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "USER AWARENESS AND ACKNOWLEDGMENT TRACKING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Department", 22),
        ("Total_Users", 14),
        ("Acknowledged", 14),
        ("Pending", 14),
        ("Acknowledgment_%", 16),
        ("Last_Campaign_Date", 18),
        ("Next_Due_Date", 16),
        ("Campaign_Method", 22),
        ("Escalation_Status", 18),
        ("Notes", 35),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate common departments
    # CUSTOMIZE: Adapt to your organization's structure
    departments = [
        "Executive Leadership",
        "Finance",
        "Human Resources",
        "IT Operations",
        "Software Development",
        "Sales",
        "Marketing",
        "Customer Support",
        "Legal",
        "Facilities",
        "External Contractors",
    ]

    # Data validations
    dv_method = DataValidation(
        type="list",
        formula1='"Email,LMS,Intranet,In-Person,Onboarding"',
        allow_blank=False
    )
    ws.add_data_validation(dv_method)

    dv_escalation = DataValidation(
        type="list",
        formula1='"None Required,Reminder Sent,Manager Notified,HR Escalation"',
        allow_blank=False
    )
    ws.add_data_validation(dv_escalation)

    # Populate rows
    for row_idx, dept in enumerate(departments, start=4):
        ws.cell(row=row_idx, column=1, value=dept)

        for c in range(2, 11):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Formula for percentage
        ws.cell(row=row_idx, column=5).value = f'=IF(B{row_idx}>0,C{row_idx}/B{row_idx}*100,"")'
        ws.cell(row=row_idx, column=5).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws.cell(row=row_idx, column=5).number_format = '0.0"%"'

        dv_method.add(ws.cell(row=row_idx, column=8))
        dv_escalation.add(ws.cell(row=row_idx, column=9))

    # Additional blank rows
    for r in range(len(departments) + 4, len(departments) + 14):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        ws.cell(row=r, column=5).value = f'=IF(B{r}>0,C{r}/B{r}*100,"")'
        ws.cell(row=r, column=5).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        dv_method.add(ws.cell(row=r, column=8))
        dv_escalation.add(ws.cell(row=r, column=9))

    # Summary row
    summary_row = len(departments) + 15
    ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=summary_row, column=2).value = f"=SUM(B4:B{summary_row-1})"
    ws.cell(row=summary_row, column=3).value = f"=SUM(C4:C{summary_row-1})"
    ws.cell(row=summary_row, column=4).value = f"=SUM(D4:D{summary_row-1})"
    ws.cell(row=summary_row, column=5).value = f'=IF(B{summary_row}>0,C{summary_row}/B{summary_row}*100,"")'
    ws.cell(row=summary_row, column=5).number_format = '0.0"%"'

    for c in range(1, 6):
        ws.cell(row=summary_row, column=c).font = Font(bold=True)
        ws.cell(row=summary_row, column=c).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws.cell(row=summary_row, column=c).border = styles["border"]

    ws.freeze_panes = "B4"


# =============================================================================
# COMMUNICATION MATRIX SHEET
# =============================================================================
def create_communication_matrix_sheet(ws, styles):
    """Create the Communication_Matrix sheet - policy dissemination assessment."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "POLICY COMMUNICATION MATRIX"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Communication_Channel", 28),
        ("Audience", 25),
        ("Frequency", 18),
        ("Last_Communication", 18),
        ("Next_Scheduled", 18),
        ("Effectiveness_Rating", 18),
        ("Owner", 22),
        ("Improvement_Actions", 35),
        ("Evidence_Ref", 18),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate communication channels
    channels = [
        ("New Employee Onboarding", "New hires", "Per hire"),
        ("Annual AUP Refresh Campaign", "All employees", "Annual"),
        ("Intranet Policy Portal", "All employees", "Always available"),
        ("Security Awareness Training", "All employees", "Annual"),
        ("Manager Briefings", "People managers", "Quarterly"),
        ("Contractor Onboarding", "New contractors", "Per engagement"),
        ("Policy Update Notifications", "All employees", "As needed"),
        ("Compliance Reminders", "Non-compliant users", "As needed"),
    ]

    # Data validations
    dv_frequency = DataValidation(
        type="list",
        formula1='"Daily,Weekly,Monthly,Quarterly,Annual,Per hire,As needed,Always available"',
        allow_blank=False
    )
    ws.add_data_validation(dv_frequency)

    dv_effectiveness = DataValidation(
        type="list",
        formula1='"Highly Effective,Effective,Needs Improvement,Ineffective"',
        allow_blank=False
    )
    ws.add_data_validation(dv_effectiveness)

    # Populate rows
    for row_idx, (channel, audience, frequency) in enumerate(channels, start=4):
        ws.cell(row=row_idx, column=1, value=channel)
        ws.cell(row=row_idx, column=2, value=audience)
        ws.cell(row=row_idx, column=3, value=frequency)

        for c in range(4, 10):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_frequency.add(ws.cell(row=row_idx, column=3))
        dv_effectiveness.add(ws.cell(row=row_idx, column=6))

    # Additional blank rows
    for r in range(len(channels) + 4, len(channels) + 14):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_frequency.add(ws.cell(row=r, column=3))
        dv_effectiveness.add(ws.cell(row=r, column=6))

    ws.freeze_panes = "B4"


# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================
def create_evidence_register_sheet(ws, styles):
    """Create the Evidence_Register sheet for audit traceability."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 45),
        ("Related_Requirement", 25),
        ("Collection_Date", 16),
        ("Location", 40),
        ("Collected_By", 25),
        ("Valid_Until", 16),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Evidence type validation
    dv_type = DataValidation(
        type="list",
        formula1='"Policy Document,Acknowledgment Record,Training Record,Screenshot,Export,Attestation,Email,Meeting Minutes"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    # Data entry rows
    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")

        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "C4"


# =============================================================================
# APPROVAL SIGN-OFF SHEET
# =============================================================================
def create_approval_signoff_sheet(ws, styles):
    """Create the Approval_SignOff sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Total Policy Requirements Assessed", "=COUNTA(Policy_Assessment!A4:A33)-COUNTBLANK(Policy_Assessment!B4:B33)"),
        ("Requirements Addressed", '=COUNTIF(Policy_Assessment!D4:D33,"Yes")'),
        ("Requirements Partially Addressed", '=COUNTIF(Policy_Assessment!D4:D33,"Partial")'),
        ("Gaps Identified", '=COUNTIF(Policy_Assessment!D4:D33,"No")'),
        ("Overall AUP Acknowledgment %", "=Awareness_Tracking!E26"),
        ("Assessment Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "" or (isinstance(value, str) and not value.startswith("=")):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    # Assessment Completed By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    completion_fields = ["Name", "Role/Title", "Department", "Email", "Date"]
    row += 1
    for field in completion_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Reviewed By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (Information Security Manager)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    review_fields = ["Name", "Date", "Signature"]
    row += 1
    for field in review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Approved By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    approval_fields = ["Name", "Date", "Approval Decision", "Signature"]
    row += 1
    for field in approval_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Decision dropdown
    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add(ws[f"B{row-2}"])

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A3"


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main() -> int:
    """
    Main execution function - orchestrates workbook creation.

    Returns:
        int: 0 on success, 1 on failure
    """
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/7] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/7] Creating Policy_Assessment sheet...")
        create_policy_assessment_sheet(wb["Policy_Assessment"], styles)

        logger.info("[3/7] Creating Asset_Coverage sheet...")
        create_asset_coverage_sheet(wb["Asset_Coverage"], styles)

        logger.info("[4/7] Creating Awareness_Tracking sheet...")
        create_awareness_tracking_sheet(wb["Awareness_Tracking"], styles)

        logger.info("[5/7] Creating Communication_Matrix sheet...")
        create_communication_matrix_sheet(wb["Communication_Matrix"], styles)

        logger.info("[6/7] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[7/7] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("Next steps:")
        logger.info("  1) Complete document information in Instructions")
        logger.info("  2) Assess policy requirements in Policy_Assessment")
        logger.info("  3) Verify asset coverage in Asset_Coverage")
        logger.info("  4) Track user acknowledgments in Awareness_Tracking")
        logger.info("  5) Review communication effectiveness")
        logger.info("  6) Link evidence in Evidence_Register")
        logger.info("  7) Obtain approvals")
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        logger.error("Install with: pip install openpyxl")
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# END OF SCRIPT
# =============================================================================
# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.10-11.1 specification
# =============================================================================
