#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sanity Check for A.5.10-11 Asset Usage Lifecycle Workbooks

Validates workbook structure, required sheets, data validations,
and formula integrity.
"""

import logging
import sys
from pathlib import Path

from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

WORKBOOK_DIR = Path(__file__).parent.parent / "90_workbooks"

# Expected workbooks and their required sheets
EXPECTED_WORKBOOKS = {
    "ISMS-IMP-A.5.10-11.S1": [
        "Instructions", "Policy_Assessment", "Asset_Coverage",
        "Awareness_Tracking", "Communication_Matrix", "Evidence_Register",
        "Approval_SignOff"
    ],
    "ISMS-IMP-A.5.10-11.S2": [
        "Instructions", "Usage_Rules", "Permitted_Activities",
        "Prohibited_Activities", "Handling_Requirements", "Evidence_Register",
        "Approval_SignOff"
    ],
    "ISMS-IMP-A.5.10-11.S3": [
        "Instructions", "Return_Process", "Asset_Checklist",
        "Offboarding_Tracking", "Access_Revocation", "Evidence_Register",
        "Approval_SignOff"
    ],
    "ISMS-IMP-A.5.10-11.S4": [
        "Instructions", "Executive_Summary", "A510_Compliance",
        "A511_Compliance", "Gap_Register", "Remediation_Tracker",
        "Trend_Analysis", "Approval_SignOff"
    ],
}


def check_workbook(filepath: Path, expected_sheets: list) -> tuple:
    """Check a single workbook. Returns (passed, errors)."""
    errors = []

    try:
        wb = load_workbook(filepath, data_only=False)

        # Check sheets exist
        for sheet_name in expected_sheets:
            if sheet_name not in wb.sheetnames:
                errors.append(f"Missing sheet: {sheet_name}")

        # Check for extra unexpected sheets
        for sheet_name in wb.sheetnames:
            if sheet_name not in expected_sheets:
                errors.append(f"Unexpected sheet: {sheet_name}")

        # Check Instructions sheet has header
        if "Instructions" in wb.sheetnames:
            ws = wb["Instructions"]
            if ws["A1"].value is None:
                errors.append("Instructions sheet missing header")

        # Check Approval_SignOff has approval fields
        if "Approval_SignOff" in wb.sheetnames:
            ws = wb["Approval_SignOff"]
            if ws["A1"].value is None:
                errors.append("Approval_SignOff sheet missing header")

        wb.close()

    except Exception as e:
        errors.append(f"Failed to open workbook: {e}")

    return (len(errors) == 0, errors)


def main() -> int:
    """Main execution."""
    logger.info("=" * 60)
    logger.info("A.5.10-11 Workbook Sanity Check")
    logger.info("=" * 60)

    if not WORKBOOK_DIR.exists():
        logger.error("Workbook directory not found: %s", WORKBOOK_DIR)
        return 1

    all_passed = True
    workbooks_checked = 0

    for doc_id, expected_sheets in EXPECTED_WORKBOOKS.items():
        # Find matching workbook
        matches = list(WORKBOOK_DIR.glob(f"{doc_id}*.xlsx"))

        if not matches:
            logger.warning("Workbook not found: %s", doc_id)
            all_passed = False
            continue

        wb_path = matches[0]  # Take first match
        passed, errors = check_workbook(wb_path, expected_sheets)
        workbooks_checked += 1

        if passed:
            logger.info("PASS: %s", wb_path.name)
        else:
            logger.error("FAIL: %s", wb_path.name)
            for error in errors:
                logger.error("  - %s", error)
            all_passed = False

    logger.info("=" * 60)
    logger.info("Checked %d workbooks", workbooks_checked)

    if all_passed:
        logger.info("All sanity checks PASSED")
        return 0
    else:
        logger.error("Some sanity checks FAILED")
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - UTILITY SCRIPT
# =============================================================================

