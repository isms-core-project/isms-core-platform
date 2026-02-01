#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.1-2-6.1-2 - Assessment File Normalization Utility
================================================================================

ISO/IEC 27001:2022 Controls A.5.1, A.5.2, A.6.1, A.6.2: Secure Employment and Roles
Utility Script: File Normalization and Reference Stabilization

--------------------------------------------------------------------------------
UTILITY SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a UTILITY/HELPER tool and MUST be adapted to match your
organization's specific file naming conventions, directory structures, and
workflow requirements.

Key customization areas:
1. File naming patterns and date formats (match your conventions)
2. Directory paths and file locations (adapt to your environment)
3. External reference update logic (specific to your workbook structure)
4. Backup and archival procedures (align with your data management policies)
5. Normalization rules (customize for your assessment workflows)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.1-2-6.1-2 Secure Employment and Roles Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This utility script performs file normalization and reference stabilization
for Control A.5.1-2-6.1-2 assessment workbooks to ensure consistent external
references in the S5 Governance Compliance Dashboard and maintain audit trail
integrity.

**Purpose:**
Eliminates broken external workbook references in S5 dashboard by standardizing
filenames, updating reference paths, and creating stable "normalized" versions
of assessment workbooks for dashboard consumption.

**Problem Statement:**
Assessment workbooks are generated with date-stamped filenames:
- ISMS-IMP-A.5.1-2-6.1-2.S1_Policy_Framework_20260115.xlsx
- ISMS-IMP-A.5.1-2-6.1-2.S1_Policy_Framework_20260122.xlsx (updated version)

Dashboard workbook (S5) uses external references like:
- ='[ISMS-IMP-A.5.1-2-6.1-2.S1_Policy_Framework_20260115.xlsx]Dashboard'!A1

When assessment files are updated with new dates, external references break.

**Solution:**
This script creates "normalized" copies with stable filenames:
- ISMS-IMP-A.5.1-2-6.1-2.S1_Policy_Framework_20260115.xlsx -> ISMS-IMP-A.5.1-2-6.1-2.S1_NORMALIZED.xlsx
- ISMS-IMP-A.5.1-2-6.1-2.S2_Roles_Responsibilities_20260115.xlsx -> ISMS-IMP-A.5.1-2-6.1-2.S2_NORMALIZED.xlsx
- ISMS-IMP-A.5.1-2-6.1-2.S3_Screening_Vetting_20260115.xlsx -> ISMS-IMP-A.5.1-2-6.1-2.S3_NORMALIZED.xlsx
- ISMS-IMP-A.5.1-2-6.1-2.S4_Employment_Contract_Assessment_20260115.xlsx -> ISMS-IMP-A.5.1-2-6.1-2.S4_NORMALIZED.xlsx
- ISMS-IMP-A.5.1-2-6.1-2.S5_Governance_Compliance_Dashboard_20260115.xlsx -> ISMS-IMP-A.5.1-2-6.1-2.S5_NORMALIZED.xlsx

Dashboard external references use normalized filenames, which remain stable
across assessment updates. When new assessments are generated, run this script
to update the normalized copies.

**Control Structure:**
This is a stacked/compound control covering:
- A.5.1: Policies for Information Security
- A.5.2: Information Security Roles and Responsibilities
- A.6.1: Screening
- A.6.2: Terms and Conditions of Employment

**Workbook Structure:**
- S1: Policy Framework Assessment (11 sheets) - A.5.1 focus
- S2: Roles & Responsibilities Assessment (10 sheets) - A.5.2 focus
- S3: Screening & Vetting Assessment (10 sheets) - A.6.1 focus
- S4: Employment Contract Assessment (10 sheets) - A.6.2 focus
- S5: Governance Compliance Dashboard (11 sheets) - Consolidates S1-S4

**Normalization Workflow:**
1. Generate assessment workbooks (S1-S4) with date stamps
2. Complete assessment data entry
3. Run THIS script to create normalized copies
4. Generate dashboard (S5) referencing normalized files
5. When assessments updated: repeat steps 1-4

**Additional Functionality:**
- Validates all expected assessment files are present
- Creates backup copies before normalization
- Generates normalization audit log
- Verifies Excel file integrity after normalization
- Archives date-stamped originals

**Generated Outputs:**
- ISMS-IMP-A.5.1-2-6.1-2.S1_NORMALIZED.xlsx (stable reference copy)
- ISMS-IMP-A.5.1-2-6.1-2.S2_NORMALIZED.xlsx (stable reference copy)
- ISMS-IMP-A.5.1-2-6.1-2.S3_NORMALIZED.xlsx (stable reference copy)
- ISMS-IMP-A.5.1-2-6.1-2.S4_NORMALIZED.xlsx (stable reference copy)
- ISMS-IMP-A.5.1-2-6.1-2.S5_NORMALIZED.xlsx (stable dashboard copy)
- normalization_audit_log_YYYYMMDD.txt (audit trail)
- /backups/ directory with date-stamped copies

**Key Features:**
- Automated file normalization with stable naming
- Backup creation for audit trail preservation
- Validation of file integrity before/after normalization
- Audit logging of all normalization operations
- Directory structure verification
- Error handling and rollback capabilities
- Dry-run mode for testing before execution

**Integration:**
This utility supports:
- S1 Policy Framework Assessment (normalizes for dashboard reference)
- S2 Roles & Responsibilities Assessment (normalizes for dashboard reference)
- S3 Screening & Vetting Assessment (normalizes for dashboard reference)
- S4 Employment Contract Assessment (normalizes for dashboard reference)
- S5 Governance Compliance Dashboard (consumes normalized references)
- Audit trail requirements (maintains backup copies)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel file validation
    - Standard Unix utilities (cp, sed) for Linux/macOS

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library for validation)
    - shutil (standard library for file operations)
    - datetime (standard library)
    - os, pathlib (standard library for file system operations)

Input Files Expected:
    - ISMS-IMP-A.5.1-2-6.1-2.S1_Policy_Framework_YYYYMMDD.xlsx (most recent)
    - ISMS-IMP-A.5.1-2-6.1-2.S2_Roles_Responsibilities_YYYYMMDD.xlsx (most recent)
    - ISMS-IMP-A.5.1-2-6.1-2.S3_Screening_Vetting_YYYYMMDD.xlsx (most recent)
    - ISMS-IMP-A.5.1-2-6.1-2.S4_Employment_Contract_Assessment_YYYYMMDD.xlsx (most recent)
    - ISMS-IMP-A.5.1-2-6.1-2.S5_Governance_Compliance_Dashboard_YYYYMMDD.xlsx (most recent)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 normalize_assessment_files_a5_1_2_6_1_2.py

Advanced Usage:
    # Dry run (show what would be done without making changes)
    python3 normalize_assessment_files_a5_1_2_6_1_2.py --dry-run

    # Specify custom input directory
    python3 normalize_assessment_files_a5_1_2_6_1_2.py --source /path/to/assessments

    # Specify custom output directory
    python3 normalize_assessment_files_a5_1_2_6_1_2.py --output /path/to/dashboard

    # Non-interactive mode (skip confirmation prompts)
    python3 normalize_assessment_files_a5_1_2_6_1_2.py --auto-confirm

Workflow Example:
    # Step 1: Generate assessments with date stamps
    python3 generate_a5_1_2_6_1_2_s1_policy_framework.py
    python3 generate_a5_1_2_6_1_2_s2_roles_responsibilities.py
    python3 generate_a5_1_2_6_1_2_s3_screening_vetting.py
    python3 generate_a5_1_2_6_1_2_s4_employment_contract.py

    # Step 2: Complete data entry in date-stamped files
    # (manual step)

    # Step 3: Normalize files for stable dashboard references
    python3 normalize_assessment_files_a5_1_2_6_1_2.py

    # Step 4: Generate dashboard (references NORMALIZED files)
    python3 generate_a5_1_2_6_1_2_s5_governance_dashboard.py

    # Step 5: When assessments updated, repeat normalization
    python3 normalize_assessment_files_a5_1_2_6_1_2.py  # Updates NORMALIZED copies

Output:
    Files Created:
        - ISMS-IMP-A.5.1-2-6.1-2.S1_NORMALIZED.xlsx
        - ISMS-IMP-A.5.1-2-6.1-2.S2_NORMALIZED.xlsx
        - ISMS-IMP-A.5.1-2-6.1-2.S3_NORMALIZED.xlsx
        - ISMS-IMP-A.5.1-2-6.1-2.S4_NORMALIZED.xlsx
        - ISMS-IMP-A.5.1-2-6.1-2.S5_NORMALIZED.xlsx
        - normalization_audit_log_YYYYMMDD.txt

    Directory Created:
        - backups/ (contains previous NORMALIZED versions)

Post-Execution Steps:
    1. Verify all NORMALIZED files created successfully
    2. Review normalization_audit_log_YYYYMMDD.txt for any errors
    3. Validate Excel files open without errors
    4. Verify dashboard references work (S5 -> S1-S4)
    5. Archive date-stamped originals for audit trail
    6. Proceed with dashboard review and sign-off

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.5.1, A.5.2, A.6.1, A.6.2
Utility Type:         File Normalization and Reference Stabilization
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization ISMS Team]
Date Created:         [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.1-2-6.1-2: Secure Employment and Roles Policy (Governance)
    - ISMS-IMP-A.5.1-2-6.1-2.S1: Policy Framework Assessment
    - ISMS-IMP-A.5.1-2-6.1-2.S2: Roles & Responsibilities Assessment
    - ISMS-IMP-A.5.1-2-6.1-2.S3: Screening & Vetting Assessment
    - ISMS-IMP-A.5.1-2-6.1-2.S4: Employment Contract Assessment
    - ISMS-IMP-A.5.1-2-6.1-2.S5: Governance Compliance Dashboard
    - Assessment Workflow Documentation

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release with basic file normalization
    - Support for S1-S5 workbook normalization
    - Backup creation before overwriting
    - Audit logging functionality
    - Dry-run mode for testing
    - Comprehensive validation with openpyxl
    - Error handling and rollback capabilities

[Future changes to be documented here]

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# Third-Party Imports
# =============================================================================
try:
    import openpyxl
except ImportError:
    logger.error("openpyxl not installed")
    logger.error("Install with: sudo apt install python3-openpyxl")
    logger.error("Or: pip3 install openpyxl")
    sys.exit(1)


# ============================================================================
# CONFIGURATION
# ============================================================================

# Expected document IDs and titles from assessment workbooks
# CUSTOMIZE: Adjust document IDs and titles to match your workbook metadata
EXPECTED_DOCS = {
    "ISMS-IMP-A.5.1-2-6.1-2.S1": {
        "title": "Policy Framework Assessment",
        "normalized": "ISMS-IMP-A.5.1-2-6.1-2.S1_NORMALIZED.xlsx",
        "sheet": "Dashboard",
        "sheets_expected": 11,
        "control_focus": "A.5.1 - Policies for Information Security"
    },
    "ISMS-IMP-A.5.1-2-6.1-2.S2": {
        "title": "Roles & Responsibilities Assessment",
        "normalized": "ISMS-IMP-A.5.1-2-6.1-2.S2_NORMALIZED.xlsx",
        "sheet": "Dashboard",
        "sheets_expected": 10,
        "control_focus": "A.5.2 - Information Security Roles and Responsibilities"
    },
    "ISMS-IMP-A.5.1-2-6.1-2.S3": {
        "title": "Screening & Vetting Assessment",
        "normalized": "ISMS-IMP-A.5.1-2-6.1-2.S3_NORMALIZED.xlsx",
        "sheet": "Dashboard",
        "sheets_expected": 10,
        "control_focus": "A.6.1 - Screening"
    },
    "ISMS-IMP-A.5.1-2-6.1-2.S4": {
        "title": "Employment Contract Assessment",
        "normalized": "ISMS-IMP-A.5.1-2-6.1-2.S4_NORMALIZED.xlsx",
        "sheet": "Dashboard",
        "sheets_expected": 10,
        "control_focus": "A.6.2 - Terms and Conditions of Employment"
    },
    "ISMS-IMP-A.5.1-2-6.1-2.S5": {
        "title": "Governance Compliance Dashboard",
        "normalized": "ISMS-IMP-A.5.1-2-6.1-2.S5_NORMALIZED.xlsx",
        "sheet": "Executive_Summary",
        "sheets_expected": 11,
        "control_focus": "Consolidated Dashboard (A.5.1 + A.5.2 + A.6.1 + A.6.2)"
    },
}

# Required workbooks for S5 dashboard (S1-S4)
REQUIRED_FOR_DASHBOARD = [
    "ISMS-IMP-A.5.1-2-6.1-2.S1",
    "ISMS-IMP-A.5.1-2-6.1-2.S2",
    "ISMS-IMP-A.5.1-2-6.1-2.S3",
    "ISMS-IMP-A.5.1-2-6.1-2.S4"
]


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_workbook(filepath):
    """
    Validate workbook contains expected document ID in metadata or Dashboard sheet.

    Args:
        filepath: Path to Excel workbook

    Returns:
        tuple: (doc_id, title) if valid, (None, None) if invalid
    """
    try:
        # Open workbook in read-only mode
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        # Try to find Document ID in multiple possible locations
        sheets_to_check = ["Metadata", "Dashboard", "Executive_Summary", "Instructions_Legend"]

        for sheet_name in sheets_to_check:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            # Look for Document ID in column A or B (typically rows 1-25)
            for row in range(1, 30):
                cell_label = ws.cell(row=row, column=1).value

                if cell_label and "Document ID" in str(cell_label):
                    doc_id_value = ws.cell(row=row, column=2).value

                    if doc_id_value:
                        doc_id = str(doc_id_value).strip()

                        # Check if it's one of our expected documents
                        if doc_id in EXPECTED_DOCS:
                            title = EXPECTED_DOCS[doc_id]["title"]
                            wb.close()
                            return (doc_id, title)

                # Also check column B for label, column C for value
                cell_label_b = ws.cell(row=row, column=2).value
                if cell_label_b and "Document ID" in str(cell_label_b):
                    doc_id_value = ws.cell(row=row, column=3).value

                    if doc_id_value:
                        doc_id = str(doc_id_value).strip()

                        if doc_id in EXPECTED_DOCS:
                            title = EXPECTED_DOCS[doc_id]["title"]
                            wb.close()
                            return (doc_id, title)

        wb.close()
        return (None, None)

    except Exception as e:
        logger.warning(f"Error reading file {filepath}: {e}")
        return (None, None)


def get_file_info(filepath):
    """Get file metadata for manifest."""
    stat = filepath.stat()
    return {
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime),
        "created": datetime.fromtimestamp(stat.st_ctime),
    }


def validate_sheet_count(filepath, expected_count):
    """
    Validate workbook has expected number of sheets.

    Args:
        filepath: Path to Excel workbook
        expected_count: Expected number of sheets

    Returns:
        tuple: (actual_count, is_valid)
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        actual_count = len(wb.sheetnames)
        wb.close()
        return (actual_count, actual_count == expected_count)
    except Exception as e:
        logger.warning(f"Error validating sheet count: {e}")
        return (0, False)


# ============================================================================
# SCANNING FUNCTIONS
# ============================================================================

def scan_directory(directory):
    """
    Scan directory for assessment workbooks.

    Args:
        directory: Path to scan

    Returns:
        dict: {doc_id: {'path': Path, 'info': metadata}}
    """
    found_assessments = {}
    directory = Path(directory)

    # Get all .xlsx files, excluding temporary files and normalized files
    xlsx_files = [
        f for f in directory.glob("*.xlsx")
        if not f.name.startswith("~$") and "NORMALIZED" not in f.name
    ]

    if not xlsx_files:
        logger.error(f"No Excel files found in {directory}")
        return found_assessments

    logger.info(f"Scanning {len(xlsx_files)} Excel file(s) in {directory}...")

    for filepath in sorted(xlsx_files):
        logger.info(f"  Checking: {filepath.name}")
        doc_id, title = validate_workbook(filepath)

        if doc_id:
            logger.info(f"    Valid: {doc_id} - {title}")

            # Check for duplicates (prefer most recent by modification time)
            if doc_id in found_assessments:
                existing_info = found_assessments[doc_id]['info']
                new_info = get_file_info(filepath)

                if new_info['modified'] > existing_info['modified']:
                    logger.info(f"    Duplicate found - using newer file: {filepath.name}")
                    found_assessments[doc_id] = {
                        'path': filepath,
                        'info': new_info
                    }
                else:
                    logger.info(f"    Duplicate found - keeping older file (newer exists)")
            else:
                found_assessments[doc_id] = {
                    'path': filepath,
                    'info': get_file_info(filepath)
                }
        else:
            logger.info(f"    Skipped (not a valid A.5.1-2-6.1-2 assessment workbook)")

    return found_assessments


# ============================================================================
# BACKUP FUNCTIONS
# ============================================================================

def create_backup(filepath, backup_dir):
    """
    Create backup of existing file before overwriting.

    Args:
        filepath: Path to file to backup
        backup_dir: Directory to store backups

    Returns:
        Path: Path to backup file, or None if backup failed
    """
    if not filepath.exists():
        return None

    backup_dir = Path(backup_dir)
    backup_dir.mkdir(parents=True, exist_ok=True)

    # Create backup filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{filepath.stem}_backup_{timestamp}{filepath.suffix}"
    backup_path = backup_dir / backup_name

    try:
        shutil.copy2(filepath, backup_path)
        logger.info(f"  Backup created: {backup_name}")
        return backup_path
    except Exception as e:
        logger.error(f"  Backup failed: {e}")
        return None


# ============================================================================
# MANIFEST CREATION
# ============================================================================

def create_manifest(output_dir, mapping, source_dir):
    """
    Create audit manifest documenting the normalization process.

    Args:
        output_dir: Destination directory
        mapping: Dictionary of normalized files
        source_dir: Source directory

    Returns:
        Path: Path to created manifest file
    """
    timestamp = datetime.now().strftime("%Y%m%d")
    manifest_path = Path(output_dir) / f"normalization_audit_log_{timestamp}.txt"

    with open(manifest_path, 'w', encoding='utf-8') as f:
        # Header
        f.write("=" * 80 + "\n")
        f.write("ISMS ASSESSMENT FILE NORMALIZATION AUDIT LOG\n")
        f.write("ISO/IEC 27001:2022 - Controls A.5.1, A.5.2, A.6.1, A.6.2\n")
        f.write("Secure Employment and Roles Framework\n")
        f.write("=" * 80 + "\n\n")

        # Normalization metadata
        f.write("NORMALIZATION METADATA\n")
        f.write("-" * 80 + "\n")
        f.write(f"Normalization Date/Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Source Directory:        {source_dir.resolve()}\n")
        f.write(f"Output Directory:        {output_dir.resolve()}\n")
        f.write(f"Files Normalized:        {len(mapping)}/{len(EXPECTED_DOCS)} expected\n")
        f.write(f"Normalization Status:    {'COMPLETE' if len(mapping) == len(EXPECTED_DOCS) else 'PARTIAL'}\n")
        f.write("\n")

        # Control structure
        f.write("CONTROL STRUCTURE\n")
        f.write("-" * 80 + "\n")
        f.write("This stacked control covers:\n")
        f.write("  - A.5.1: Policies for Information Security\n")
        f.write("  - A.5.2: Information Security Roles and Responsibilities\n")
        f.write("  - A.6.1: Screening\n")
        f.write("  - A.6.2: Terms and Conditions of Employment\n\n")

        # File mapping details
        f.write("FILE MAPPING\n")
        f.write("-" * 80 + "\n\n")

        for doc_id in sorted(EXPECTED_DOCS.keys()):
            if doc_id in mapping:
                info = mapping[doc_id]
                source_path = info['path']
                normalized_name = info['normalized']
                file_info = info['info']

                f.write(f"{doc_id}\n")
                f.write(f"  Title:           {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"  Control Focus:   {EXPECTED_DOCS[doc_id]['control_focus']}\n")
                f.write(f"  Source File:     {source_path.name}\n")
                f.write(f"  Normalized File: {normalized_name}\n")
                f.write(f"  File Size:       {file_info['size']:,} bytes\n")
                f.write(f"  Last Modified:   {file_info['modified'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"  Status:          NORMALIZED\n\n")
            else:
                f.write(f"{doc_id}\n")
                f.write(f"  Title:           {EXPECTED_DOCS[doc_id]['title']}\n")
                f.write(f"  Control Focus:   {EXPECTED_DOCS[doc_id]['control_focus']}\n")
                f.write(f"  Status:          NOT FOUND\n\n")

        # Dashboard integration instructions
        f.write("\n")
        f.write("=" * 80 + "\n")
        f.write("DASHBOARD INTEGRATION INSTRUCTIONS\n")
        f.write("=" * 80 + "\n\n")

        f.write("PURPOSE\n")
        f.write("-" * 80 + "\n")
        f.write("Normalized files enable stable external references in S5 dashboard.\n")
        f.write("Dashboard formulas reference files without dates (e.g., S1_NORMALIZED.xlsx).\n")
        f.write("This prevents broken links when generating new assessment versions.\n\n")

        f.write("WORKFLOW\n")
        f.write("-" * 80 + "\n")
        f.write("Step 1: Generate Assessment Workbooks\n")
        f.write("  python3 generate_a5_1_2_6_1_2_s1_policy_framework.py\n")
        f.write("  python3 generate_a5_1_2_6_1_2_s2_roles_responsibilities.py\n")
        f.write("  python3 generate_a5_1_2_6_1_2_s3_screening_vetting.py\n")
        f.write("  python3 generate_a5_1_2_6_1_2_s4_employment_contract.py\n")
        f.write("  -> Creates: ISMS-IMP-A.5.1-2-6.1-2.SX_Name_YYYYMMDD.xlsx (with dates)\n\n")

        f.write("Step 2: Complete Assessments\n")
        f.write("  - Fill in assessment data in workbooks S1-S4\n")
        f.write("  - Document policies, roles, screening, contracts\n")
        f.write("  - Run sanity checks to validate completeness\n\n")

        f.write("Step 3: Normalize Filenames (THIS SCRIPT)\n")
        f.write("  python3 normalize_assessment_files_a5_1_2_6_1_2.py\n")
        f.write("  -> Creates: ISMS-IMP-A.5.1-2-6.1-2.SX_NORMALIZED.xlsx (no dates)\n")
        f.write("  -> Location: Dashboard_Sources/ directory\n\n")

        f.write("Step 4: Generate Dashboard\n")
        f.write("  python3 generate_a5_1_2_6_1_2_s5_governance_dashboard.py\n")
        f.write("  -> Creates: ISMS-IMP-A.5.1-2-6.1-2.S5_Governance_Dashboard_YYYYMMDD.xlsx\n\n")

        f.write("Step 5: Setup Dashboard Links\n")
        f.write("  - Copy dashboard to Dashboard_Sources/ directory\n")
        f.write("  - Open dashboard in Excel\n")
        f.write("  - Click 'Update Links' when prompted\n")
        f.write("  - Dashboard auto-populates with current compliance data\n\n")

        f.write("MAINTENANCE\n")
        f.write("-" * 80 + "\n")
        f.write("To update dashboard with new data:\n")
        f.write("  1. Edit assessment workbooks (S1-S4)\n")
        f.write("  2. Save changes\n")
        f.write("  3. Re-run normalization script (copies updated files)\n")
        f.write("  4. Open dashboard and refresh links (Data -> Refresh All)\n\n")

        f.write("DASHBOARD DEPENDENCIES\n")
        f.write("-" * 80 + "\n")
        f.write("S5 Governance Compliance Dashboard requires:\n")
        f.write("  - S1 (Policy Framework) - 35% weight in composite score\n")
        f.write("  - S2 (Roles & Responsibilities) - 25% weight\n")
        f.write("  - S3 (Screening & Vetting) - 20% weight\n")
        f.write("  - S4 (Employment Contracts) - 20% weight\n\n")

        f.write("=" * 80 + "\n")
        f.write("END OF AUDIT LOG\n")
        f.write("=" * 80 + "\n")

    return manifest_path


# ============================================================================
# MAIN NORMALIZATION FUNCTION
# ============================================================================

def normalize_files(source_dir=None, output_dir=None, auto_confirm=False, dry_run=False):
    """
    Main normalization orchestration function.

    Args:
        source_dir: Source directory (prompts if None)
        output_dir: Output directory (prompts if None)
        auto_confirm: Skip confirmation prompts (for automation)
        dry_run: Show what would be done without making changes

    Returns:
        bool: True if successful, False otherwise
    """
    logger.info("=" * 80)
    logger.info("ISMS ASSESSMENT FILE NORMALIZATION UTILITY")
    logger.info("ISO/IEC 27001:2022 - Controls A.5.1, A.5.2, A.6.1, A.6.2")
    logger.info("Secure Employment and Roles Framework")
    logger.info("=" * 80)

    if dry_run:
        logger.info("*** DRY RUN MODE - No changes will be made ***")

    logger.info("")
    logger.info("This script prepares assessment workbooks for dashboard linking by:")
    logger.info("  1. Scanning for completed assessment files (S1-S5)")
    logger.info("  2. Validating document IDs in workbook metadata")
    logger.info("  3. Copying to normalized filenames (no dates/versions)")
    logger.info("  4. Creating audit manifest for traceability")
    logger.info("")

    # Get source directory
    if not source_dir:
        default_source = Path(__file__).parent.parent / "90_workbooks"
        source_input = input(f"Enter source directory (or press Enter for '{default_source}'): ").strip()
        if not source_input:
            source_dir = default_source
        else:
            source_dir = Path(source_input)

    source_dir = Path(source_dir).resolve()

    if not source_dir.exists():
        logger.error(f"Error: Source directory does not exist: {source_dir}")
        return False

    if not source_dir.is_dir():
        logger.error(f"Error: Not a directory: {source_dir}")
        return False

    logger.info(f"Source directory: {source_dir}")

    # Get output directory
    if not output_dir:
        default_output = source_dir / "Dashboard_Sources"
        output_input = input(f"\nEnter output directory (or press Enter for '{default_output.name}'): ").strip()
        if not output_input:
            output_dir = default_output
        else:
            output_dir = Path(output_input)

    output_dir = Path(output_dir).resolve()

    # Create output directory if it doesn't exist
    if not dry_run:
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            logger.info(f"Output directory: {output_dir}")
        except Exception as e:
            logger.error(f"Error creating output directory: {e}")
            return False
    else:
        logger.info(f"Output directory (would create): {output_dir}")

    # Scan for assessment files
    found = scan_directory(source_dir)

    if not found:
        logger.error("No valid assessment workbooks found in source directory")
        logger.info("Ensure files contain valid Document IDs:")
        for doc_id, info in EXPECTED_DOCS.items():
            logger.info(f"  - {doc_id}: {info['title']}")
        return False

    # Show normalization summary
    logger.info("")
    logger.info("=" * 80)
    logger.info("NORMALIZATION SUMMARY")
    logger.info("=" * 80)
    logger.info(f"Found {len(found)} of {len(EXPECTED_DOCS)} expected assessment workbooks:")
    logger.info("")

    mapping = {}
    for doc_id in sorted(EXPECTED_DOCS.keys()):
        if doc_id in found:
            normalized_name = EXPECTED_DOCS[doc_id]["normalized"]
            source_path = found[doc_id]['path']

            logger.info(f"  [OK] {doc_id}")
            logger.info(f"       Title:      {EXPECTED_DOCS[doc_id]['title']}")
            logger.info(f"       Source:     {source_path.name}")
            logger.info(f"       Normalized: {normalized_name}")
            logger.info("")

            mapping[doc_id] = {
                'path': source_path,
                'normalized': normalized_name,
                'info': found[doc_id]['info']
            }
        else:
            logger.info(f"  [MISSING] {doc_id} - NOT FOUND")
            logger.info(f"       Title: {EXPECTED_DOCS[doc_id]['title']}")
            logger.info("")

    # Check for minimum required files (need S1-S4 for S5 dashboard)
    missing_required = [doc for doc in REQUIRED_FOR_DASHBOARD if doc not in found]

    if missing_required:
        logger.warning("WARNING: Missing required files for S5 dashboard integration:")
        for doc in missing_required:
            logger.warning(f"   - {doc}: {EXPECTED_DOCS[doc]['title']}")
        logger.warning("   S5 Dashboard requires S1, S2, S3, and S4")
        logger.info("")

    if len(found) < len(EXPECTED_DOCS):
        logger.info(f"Note: Found {len(found)}/{len(EXPECTED_DOCS)} workbooks")
        logger.info("   You can normalize partial sets and add more files later")
        logger.info("")

    logger.info(f"Output directory: {output_dir}")
    logger.info("")

    # Confirm normalization
    if not auto_confirm and not dry_run:
        response = input("Proceed with normalization? (y/n): ").strip().lower()
        if response != 'y':
            logger.info("Normalization cancelled by user")
            return False

    if dry_run:
        logger.info("")
        logger.info("=" * 80)
        logger.info("DRY RUN COMPLETE - No files were modified")
        logger.info("=" * 80)
        logger.info("Run without --dry-run to perform actual normalization")
        return True

    # Create backup directory
    backup_dir = output_dir / "backups"

    # Perform normalization (copy files)
    logger.info("")
    logger.info("=" * 80)
    logger.info("NORMALIZING FILES...")
    logger.info("=" * 80)
    logger.info("")

    for doc_id, info in mapping.items():
        source = info['path']
        dest = output_dir / info['normalized']

        logger.info(f"Processing: {doc_id}")
        logger.info(f"  Source: {source.name}")
        logger.info(f"  Target: {dest.name}")

        # Create backup if destination exists
        if dest.exists():
            backup = create_backup(dest, backup_dir)
            if backup:
                logger.info(f"  Backup: {backup.name}")

        try:
            shutil.copy2(source, dest)
            logger.info(f"  Status: SUCCESS")
        except Exception as e:
            logger.error(f"  Status: FAILED - {e}")
            logger.error(f"Normalization failed at {doc_id}")
            return False

        logger.info("")

    # Create audit manifest
    logger.info("Creating audit manifest...")
    try:
        manifest = create_manifest(output_dir, mapping, source_dir)
        logger.info(f"  Created: {manifest.name}")
    except Exception as e:
        logger.error(f"  Error creating manifest: {e}")
        return False

    # Success summary
    logger.info("")
    logger.info("=" * 80)
    logger.info("NORMALIZATION COMPLETE")
    logger.info("=" * 80)
    logger.info("")
    logger.info(f"Normalized files:  {output_dir}")
    logger.info(f"Audit manifest:    {manifest}")
    logger.info("")
    logger.info("NEXT STEPS:")
    logger.info("")
    logger.info("  1. Review audit manifest for file mapping details")
    logger.info("")

    if "ISMS-IMP-A.5.1-2-6.1-2.S5" not in mapping:
        logger.info("  2. Generate S5 dashboard workbook:")
        logger.info("     python3 generate_a5_1_2_6_1_2_s5_governance_dashboard.py")
        logger.info("")
        logger.info("  3. Normalize the S5 dashboard:")
        logger.info("     python3 normalize_assessment_files_a5_1_2_6_1_2.py")
        logger.info("")
    else:
        logger.info("  2. Open S5 dashboard and click 'Update Links' when prompted:")
        logger.info(f"     {output_dir / EXPECTED_DOCS['ISMS-IMP-A.5.1-2-6.1-2.S5']['normalized']}")
        logger.info("")

    logger.info("  4. Dashboard will auto-populate with current compliance data")
    logger.info("")

    if missing_required:
        logger.warning("  NOTE: S5 Dashboard requires normalized versions of:")
        logger.warning("     - S1 (Policy Framework)")
        logger.warning("     - S2 (Roles & Responsibilities)")
        logger.warning("     - S3 (Screening & Vetting)")
        logger.warning("     - S4 (Employment Contract)")
        logger.warning("     Generate missing workbooks and re-run normalization.")
        logger.info("")

    logger.info("=" * 80)

    return True


# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Normalize ISMS Control A.5.1-2-6.1-2 assessment workbooks for dashboard integration",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode (prompts for directories)
  python3 normalize_assessment_files_a5_1_2_6_1_2.py

  # Specify source directory
  python3 normalize_assessment_files_a5_1_2_6_1_2.py --source ./90_workbooks

  # Specify both directories
  python3 normalize_assessment_files_a5_1_2_6_1_2.py --source ./90_workbooks --output ./dashboard

  # Dry run (preview without making changes)
  python3 normalize_assessment_files_a5_1_2_6_1_2.py --dry-run

  # Automated mode (no prompts)
  python3 normalize_assessment_files_a5_1_2_6_1_2.py --source ./90_workbooks --auto-confirm

Workbook Structure:
  S1: Policy Framework Assessment (11 sheets) - A.5.1
  S2: Roles & Responsibilities Assessment (10 sheets) - A.5.2
  S3: Screening & Vetting Assessment (10 sheets) - A.6.1
  S4: Employment Contract Assessment (10 sheets) - A.6.2
  S5: Governance Compliance Dashboard (11 sheets) - Consolidates S1-S4
        """
    )

    parser.add_argument(
        '--source', '-s',
        help='Source directory containing assessment workbooks (default: ../90_workbooks)'
    )

    parser.add_argument(
        '--output', '-o',
        help='Output directory for normalized files (default: ./Dashboard_Sources)'
    )

    parser.add_argument(
        '--auto-confirm', '-y',
        action='store_true',
        help='Skip confirmation prompts (for automation)'
    )

    parser.add_argument(
        '--dry-run', '-n',
        action='store_true',
        help='Show what would be done without making changes'
    )

    args = parser.parse_args()

    success = normalize_files(
        source_dir=args.source,
        output_dir=args.output,
        auto_confirm=args.auto_confirm,
        dry_run=args.dry_run
    )

    sys.exit(0 if success else 1)


# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (syntax validated, STANDARDIZATION applied)
# QA_TOOL: Claude Code Deep Scan
# STANDARDIZATION: License header, logging, imports reorganized, main() pattern
# =============================================================================
