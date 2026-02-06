#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.12-13.S2 - Labelling Procedures and Standards Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.13: Labelling of Information
"An appropriate set of procedures for information labelling should be developed
and implemented in accordance with the information classification scheme
adopted by the organization."

Assessment Domain 2 of 4: Labelling Procedures and Standards

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel workbook for defining and managing
information labelling procedures, standards, and implementation guidelines.

**Purpose:**
Establish consistent labelling practices that communicate classification levels
effectively across all information assets (physical and digital).

**Generated Workbook Structure:**
1. Instructions - Guidance on labelling procedures
2. Labelling_Standards - Visual labels and format specifications
3. Digital_Labelling - Electronic document and metadata standards
4. Physical_Labelling - Paper and media labelling requirements
5. Automation_Tools - Automated labelling solutions
6. Evidence_Register - Audit evidence tracking
7. Approval_SignOff - Procedures approval

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.12-13.S2"
WORKBOOK_NAME = "Labelling Procedures and Standards"
CONTROL_ID = "A.5.13"
CONTROL_NAME = "Labelling of Information"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "level_restricted": {
            "fill": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            "font": Font(bold=True, color="FFFFFF"),
        },
        "level_confidential": {
            "fill": PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid"),
            "font": Font(bold=True),
        },
        "level_internal": {
            "fill": PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid"),
            "font": Font(bold=True),
        },
        "level_public": {
            "fill": PatternFill(start_color="74C0FC", end_color="74C0FC", fill_type="solid"),
            "font": Font(bold=True),
        },
    }


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    for attr, value in style_dict.items():
        if attr != "border":
            setattr(cell, attr, value)


def set_column_widths(ws, widths):
    """Set column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    apply_style(ws["A2"], styles["subheader"])

    instructions = [
        "",
        "PURPOSE",
        "This workbook defines the organization's information labelling procedures as required by",
        "ISO 27001:2022 Control A.5.13. Labelling ensures that classification is clearly communicated",
        "through visual markers, metadata, and physical identifiers.",
        "",
        "LABELLING PRINCIPLES",
        "• Labels must reflect the classification scheme established under A.5.12",
        "• Labels should be easily recognizable and consistently applied",
        "• Both physical and digital information must be appropriately labelled",
        "• Metadata should be used for automated processing and management",
        "• Labelling procedures must address internal and external transmission",
        "",
        "LABELLING METHODS",
        "• Document headers and footers",
        "• Watermarks (visible or invisible)",
        "• Metadata tags and properties",
        "• Physical stamps and stickers",
        "• Color-coded indicators",
        "• Banners and visual indicators",
        "",
        "HOW TO USE THIS WORKBOOK",
        "1. Define visual label standards in Labelling_Standards sheet",
        "2. Specify digital labelling requirements in Digital_Labelling sheet",
        "3. Document physical labelling procedures in Physical_Labelling sheet",
        "4. Record automation tools and solutions in Automation_Tools sheet",
        "5. Track evidence in Evidence_Register",
        "6. Obtain approval in Approval_SignOff",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, line in enumerate(instructions, start=4):
        ws[f"A{i}"] = line
        if line in ["PURPOSE", "LABELLING PRINCIPLES", "LABELLING METHODS", "HOW TO USE THIS WORKBOOK"]:
            ws[f"A{i}"].font = Font(bold=True, size=11)

    set_column_widths(ws, {"A": 100})
    logger.info("Created Instructions sheet")


def create_labelling_standards_sheet(ws, styles):
    """Create the Labelling Standards sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Visual Labelling Standards"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Define visual indicators, colors, and formats for each classification level"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Classification Level", "Display Text", "Color (Hex)",
        "Header Format", "Footer Format", "Watermark Text",
        "Banner Style", "Icon/Symbol", "Font Requirements"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    standards = [
        ["RESTRICTED", "🔴 RESTRICTED", "#FF6B6B",
         "RESTRICTED - [Document Title]", "Classification: RESTRICTED | Page X of Y | [Date]",
         "RESTRICTED - DO NOT DISTRIBUTE", "Red banner, top and bottom",
         "🔴 (red circle) or padlock", "Bold, uppercase, red text"],
        ["CONFIDENTIAL", "🟠 CONFIDENTIAL", "#FFA94D",
         "CONFIDENTIAL - [Document Title]", "Classification: CONFIDENTIAL | Page X of Y",
         "CONFIDENTIAL", "Orange banner, top only",
         "🟠 (orange circle)", "Bold, uppercase, orange text"],
        ["INTERNAL", "🟢 INTERNAL", "#69DB7C",
         "[Document Title]", "Internal Use Only | Page X of Y",
         "INTERNAL USE ONLY", "Green footer line",
         "🟢 (green circle)", "Normal, green text footer"],
        ["PUBLIC", "🔵 PUBLIC", "#74C0FC",
         "[Document Title]", "Page X of Y",
         "None required", "Optional blue indicator",
         "🔵 (blue circle) optional", "Standard formatting"],
    ]

    for row_idx, standard in enumerate(standards, start=5):
        for col_idx, value in enumerate(standard, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        # Apply level colors to first column
        level = standard[0].lower()
        if level == "restricted":
            apply_style(ws.cell(row=row_idx, column=1), styles["level_restricted"])
        elif level == "confidential":
            apply_style(ws.cell(row=row_idx, column=1), styles["level_confidential"])
        elif level == "internal":
            apply_style(ws.cell(row=row_idx, column=1), styles["level_internal"])
        elif level == "public":
            apply_style(ws.cell(row=row_idx, column=1), styles["level_public"])

    set_column_widths(ws, {
        "A": 18, "B": 18, "C": 12, "D": 30, "E": 40,
        "F": 25, "G": 25, "H": 20, "I": 25
    })
    logger.info("Created Labelling_Standards sheet")


def create_digital_labelling_sheet(ws, styles):
    """Create the Digital Labelling sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Digital Information Labelling Requirements"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = "Metadata, properties, and electronic marking standards"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Asset Type", "Labelling Method", "Metadata Fields",
        "Automation", "Validation", "Responsibility", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    digital_assets = [
        ["Microsoft Office Documents", "Document properties + Header/Footer",
         "Classification, Author, Created, Modified, Keywords",
         "Microsoft Information Protection (MIP)", "DLP policy check", "Document author", ""],
        ["PDF Documents", "Document properties + Watermark",
         "Classification, Creator, Custom metadata",
         "Adobe Acrobat automation", "PDF/A validation", "Document author", ""],
        ["Emails", "X-headers + Visual banner + Subject prefix",
         "X-Classification, X-Sensitivity, X-Retention",
         "Exchange transport rules", "Mail flow rules", "Email sender", ""],
        ["SharePoint/OneDrive", "Sensitivity labels + Metadata columns",
         "Classification column, Retention label",
         "Microsoft Purview", "Retention policies", "Site owner", ""],
        ["Database Records", "Classification column + Row-level security",
         "data_classification, sensitivity_level",
         "Database triggers", "Access control checks", "Data owner", ""],
        ["Source Code", "File headers + Repository metadata",
         "Classification comment, .gitattributes",
         "Pre-commit hooks", "Code review", "Developer", ""],
        ["Images/Media", "EXIF/XMP metadata + Visible watermark",
         "Classification, Copyright, Creator",
         "DAM system", "Metadata extraction", "Content owner", ""],
        ["API Responses", "Response headers + Payload metadata",
         "X-Data-Classification header",
         "API gateway rules", "Response validation", "API owner", ""],
    ]

    for row_idx, asset in enumerate(digital_assets, start=5):
        for col_idx, value in enumerate(asset, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx == 7:  # Status column
                cell.fill = styles["input_cell"]["fill"]

    dv_status = DataValidation(
        type="list",
        formula1='"✅ Implemented,⚠️ In Progress,❌ Not Implemented,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("G5:G30")

    set_column_widths(ws, {
        "A": 25, "B": 35, "C": 40, "D": 30, "E": 20, "F": 18, "G": 18
    })
    logger.info("Created Digital_Labelling sheet")


def create_physical_labelling_sheet(ws, styles):
    """Create the Physical Labelling sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Physical Information Labelling Requirements"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Paper documents, removable media, and physical asset labelling"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Asset Type", "Labelling Method", "Label Location",
        "Label Format", "Durability", "Responsible Party", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    physical_assets = [
        ["Paper Documents", "Stamp + Header/Footer printing",
         "Top right corner + Page footer",
         "Classification level + Date + Page number",
         "Indelible stamp ink", "Document creator", ""],
        ["Printed Reports", "Automatic header/footer + Cover page",
         "Every page header + Report cover",
         "Classification banner + Distribution list",
         "Print-time application", "Report generator", ""],
        ["USB Drives", "Permanent label + Engraving",
         "External surface visible",
         "Classification + Asset ID + Owner",
         "Tamper-evident label", "IT Asset Management", ""],
        ["External Hard Drives", "Asset label + Classification sticker",
         "Top surface + Visible when stored",
         "Classification + Serial + Encryption status",
         "Durable adhesive label", "IT Asset Management", ""],
        ["Backup Tapes", "Barcode label + Classification indicator",
         "Spine and front face",
         "Tape ID + Classification + Date + Retention",
         "Tape-specific labels", "Backup Administrator", ""],
        ["CD/DVD Media", "Permanent marker + Printed label",
         "Top surface (non-data side)",
         "Classification + Contents + Date",
         "CD-safe permanent marker", "Media creator", ""],
        ["File Folders", "Tab label + Cover stamp",
         "Tab edge + Front cover",
         "Classification + Contents description",
         "Adhesive label or stamp", "Records Management", ""],
        ["Storage Boxes", "Box label + Seal",
         "Front and top of box",
         "Classification + Contents + Retention + Destruction date",
         "Weatherproof label", "Records Management", ""],
    ]

    for row_idx, asset in enumerate(physical_assets, start=5):
        for col_idx, value in enumerate(asset, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx == 7:  # Status column
                cell.fill = styles["input_cell"]["fill"]

    dv_status = DataValidation(
        type="list",
        formula1='"✅ Implemented,⚠️ In Progress,❌ Not Implemented,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("G5:G30")

    set_column_widths(ws, {
        "A": 20, "B": 35, "C": 30, "D": 40, "E": 20, "F": 20, "G": 18
    })
    logger.info("Created Physical_Labelling sheet")


def create_automation_tools_sheet(ws, styles):
    """Create the Automation Tools sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Labelling Automation Tools and Solutions"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = "Automated solutions for consistent labelling enforcement"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Tool/Solution", "Vendor", "Scope/Coverage",
        "Key Features", "Integration Points", "License Type",
        "Implementation Status", "Owner"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    tools = [
        ["Microsoft Purview Information Protection", "Microsoft",
         "M365, Azure, Windows endpoints",
         "Sensitivity labels, Auto-labelling, DLP integration",
         "M365, SharePoint, Exchange, Teams",
         "M365 E5 / EMS E5", "", ""],
        ["Titus Classification Suite", "HelpSystems",
         "Cross-platform documents",
         "User-driven + automated classification, Visual marking",
         "Office, Email, File shares",
         "Per-user subscription", "", ""],
        ["Boldon James Classifier", "GRCI Group",
         "Enterprise documents and email",
         "Policy-based labelling, Watermarking, Metadata",
         "Office, Outlook, File systems",
         "Enterprise license", "", ""],
        ["Varonis Data Classification Engine", "Varonis",
         "File servers, cloud storage",
         "Content inspection, Sensitive data discovery",
         "Windows, NAS, SharePoint, Box",
         "Per-TB pricing", "", ""],
        ["Digital Guardian", "Fortra",
         "Endpoints, network, cloud",
         "Data discovery, Classification, DLP",
         "Cross-platform",
         "Per-endpoint", "", ""],
        ["Custom Scripts/Automation", "Internal",
         "Specific use cases",
         "PowerShell, Python automation",
         "APIs, file systems",
         "N/A", "", ""],
    ]

    for row_idx, tool in enumerate(tools, start=5):
        for col_idx, value in enumerate(tool, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx in [7, 8]:  # Status and Owner columns
                cell.fill = styles["input_cell"]["fill"]

    dv_status = DataValidation(
        type="list",
        formula1='"✅ Deployed,⚠️ Pilot,🔄 Evaluating,❌ Not Implemented,Planned"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("G5:G20")

    set_column_widths(ws, {
        "A": 35, "B": 15, "C": 25, "D": 45, "E": 30, "F": 18, "G": 20, "H": 15
    })
    logger.info("Created Automation_Tools sheet")


def create_evidence_register_sheet(ws, styles):
    """Create the Evidence Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Evidence Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Track audit evidence supporting labelling procedures implementation"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Evidence ID", "Description", "Evidence Type",
        "Related Procedure", "Location", "Collected Date",
        "Collected By", "Verification Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    for row in range(5, 25):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            if col in [1, 3, 7, 8]:
                cell.fill = styles["input_cell"]["fill"]

    dv_type = DataValidation(
        type="list",
        formula1='"Procedure document,Configuration screenshot,Label sample,Training record,Tool export,Audit report,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C5:C30")

    dv_status = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending Review,❌ Not Verified,Expired"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H30")

    set_column_widths(ws, {
        "A": 15, "B": 40, "C": 22, "D": 25, "E": 30, "F": 15, "G": 15, "H": 18
    })
    logger.info("Created Evidence_Register sheet")


def create_approval_sheet(ws, styles):
    """Create the Approval Sign-Off sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "Labelling Procedures Approval"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "Formal approval record for labelling procedures"
    apply_style(ws["A2"], styles["subheader"])

    ws["A4"] = "Procedure Version:"
    ws["B4"] = "1.0"
    ws["A5"] = "Effective Date:"
    ws["B5"] = ""
    ws["A6"] = "Next Review Date:"
    ws["B6"] = ""
    ws["A7"] = "Procedure Owner:"
    ws["B7"] = ""

    for row in range(4, 8):
        ws[f"A{row}"].font = Font(bold=True)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]

    ws["A10"] = "APPROVAL SIGNATURES"
    ws["A10"].font = Font(bold=True, size=12)

    approval_headers = ["Role", "Name", "Signature", "Date", "Decision", "Comments"]
    for col, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=12, column=col, value=header)
        apply_style(cell, styles["column_header"])

    approvers = [
        "Information Security Officer",
        "IT Operations Manager",
        "Records Management Lead",
        "Compliance Officer",
    ]

    for row_idx, role in enumerate(approvers, start=13):
        ws.cell(row=row_idx, column=1, value=role).border = styles["border"]
        for col in range(2, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected,Pending"',
        allow_blank=True
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add("E13:E20")

    set_column_widths(ws, {"A": 30, "B": 25, "C": 20, "D": 15, "E": 22, "F": 30})
    logger.info("Created Approval_SignOff sheet")


# =============================================================================
# MAIN FUNCTION
# =============================================================================
def main():
    """Main entry point."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS-IMP-A.5.12-13.S2 Labelling Procedures Generator")
        logger.info("=" * 70)

        wb = Workbook()
        styles = setup_styles()

        ws_instructions = wb.active
        ws_instructions.title = "Instructions"

        ws_standards = wb.create_sheet("Labelling_Standards")
        ws_digital = wb.create_sheet("Digital_Labelling")
        ws_physical = wb.create_sheet("Physical_Labelling")
        ws_automation = wb.create_sheet("Automation_Tools")
        ws_evidence = wb.create_sheet("Evidence_Register")
        ws_approval = wb.create_sheet("Approval_SignOff")

        create_instructions_sheet(ws_instructions, styles)
        create_labelling_standards_sheet(ws_standards, styles)
        create_digital_labelling_sheet(ws_digital, styles)
        create_physical_labelling_sheet(ws_physical, styles)
        create_automation_tools_sheet(ws_automation, styles)
        create_evidence_register_sheet(ws_evidence, styles)
        create_approval_sheet(ws_approval, styles)

        wb.save(OUTPUT_FILENAME)
        logger.info("=" * 70)
        logger.info("SUCCESS: Workbook saved as %s", OUTPUT_FILENAME)
        logger.info("=" * 70)
        return 0

    except Exception as e:
        logger.error("FAILED: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.12-13 Classification & Labelling
# =============================================================================
