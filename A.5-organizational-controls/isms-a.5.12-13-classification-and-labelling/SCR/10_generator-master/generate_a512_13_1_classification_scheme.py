#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.12-13.S1 - Classification Scheme Definition Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.12: Classification of Information
"Information should be classified according to the information security needs
of the organization based on confidentiality, integrity, availability and
relevant interested party requirements."

Assessment Domain 1 of 4: Classification Scheme Definition

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel workbook for defining and managing
the organization's information classification scheme.

**Purpose:**
Establish a consistent, organization-wide approach to classifying information
based on its sensitivity, value, and legal/regulatory requirements.

**Generated Workbook Structure:**
1. Instructions - Guidance on classification scheme design
2. Classification_Levels - Define classification tiers (Public to Restricted)
3. Handling_Requirements - Security controls per classification level
4. CIA_Matrix - Confidentiality, Integrity, Availability requirements
5. Regulatory_Mapping - Map classifications to regulatory requirements
6. Evidence_Register - Audit evidence tracking
7. Approval_SignOff - Scheme approval and governance

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.12-13.S1"
WORKBOOK_NAME = "Classification Scheme Definition"
CONTROL_ID = "A.5.12"
CONTROL_NAME = "Classification of Information"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "level_restricted": {
            "fill": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            "font": Font(bold=True, color="FFFFFF"),
        },
        "level_confidential": {
            "fill": PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid"),
            "font": Font(bold=True),
        },
        "level_internal": {
            "fill": PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid"),
            "font": Font(bold=True),
        },
        "level_public": {
            "fill": PatternFill(start_color="74C0FC", end_color="74C0FC", fill_type="solid"),
            "font": Font(bold=True),
        },
    }


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    for attr, value in style_dict.items():
        if attr != "border":
            setattr(cell, attr, value)


def set_column_widths(ws, widths):
    """Set column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    apply_style(ws["A2"], styles["subheader"])

    instructions = [
        "",
        "PURPOSE",
        "This workbook establishes the organization's information classification scheme as required by",
        "ISO 27001:2022 Control A.5.12. Classification enables appropriate protection of information",
        "based on its sensitivity, value, criticality, and legal/regulatory requirements.",
        "",
        "CLASSIFICATION PRINCIPLES",
        "• Information should be classified based on confidentiality, integrity, and availability needs",
        "• Classification decisions should consider legal, regulatory, and contractual requirements",
        "• The information owner is responsible for classification decisions",
        "• Classification should be reviewed periodically and when information value changes",
        "• Over-classification wastes resources; under-classification creates risk",
        "",
        "STANDARD CLASSIFICATION LEVELS",
        "Level 1 - PUBLIC: Information approved for unrestricted distribution",
        "Level 2 - INTERNAL: Information for internal use, not for public release",
        "Level 3 - CONFIDENTIAL: Sensitive business information requiring protection",
        "Level 4 - RESTRICTED: Highly sensitive information with strict access controls",
        "",
        "HOW TO USE THIS WORKBOOK",
        "1. Review and customize classification levels in Classification_Levels sheet",
        "2. Define handling requirements for each level in Handling_Requirements sheet",
        "3. Complete the CIA_Matrix to define security requirements per level",
        "4. Map classifications to regulatory requirements in Regulatory_Mapping sheet",
        "5. Document evidence in Evidence_Register",
        "6. Obtain formal approval in Approval_SignOff",
        "",
        "KEY STAKEHOLDERS",
        "• Information Security Officer - Scheme design and governance",
        "• Data Protection Officer - Regulatory alignment",
        "• Business Unit Owners - Classification decisions",
        "• IT Operations - Technical implementation",
        "• Legal/Compliance - Regulatory requirements",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, line in enumerate(instructions, start=4):
        ws[f"A{i}"] = line
        if line in ["PURPOSE", "CLASSIFICATION PRINCIPLES", "STANDARD CLASSIFICATION LEVELS",
                    "HOW TO USE THIS WORKBOOK", "KEY STAKEHOLDERS"]:
            ws[f"A{i}"].font = Font(bold=True, size=11)

    set_column_widths(ws, {"A": 100})
    logger.info("Created Instructions sheet")


def create_classification_levels_sheet(ws, styles):
    """Create the Classification Levels definition sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Classification Levels Definition"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Define organizational classification tiers and their characteristics"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Level ID", "Level Name", "Display Label", "Color Code",
        "Description", "Impact if Disclosed", "Examples",
        "Default Retention", "Review Frequency", "Owner Approval Required"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Sample classification levels
    levels = [
        ["L4", "RESTRICTED", "🔴 RESTRICTED", "#FF6B6B",
         "Highly sensitive information requiring strictest controls",
         "Severe damage to organization, legal liability, regulatory penalties",
         "PII, financial records, trade secrets, M&A data, security configs",
         "7 years", "Annual", "Yes - Executive"],
        ["L3", "CONFIDENTIAL", "🟠 CONFIDENTIAL", "#FFA94D",
         "Sensitive business information requiring protection",
         "Significant business impact, competitive disadvantage",
         "Internal financials, customer lists, contracts, HR records",
         "5 years", "Annual", "Yes - Manager"],
        ["L2", "INTERNAL", "🟢 INTERNAL", "#69DB7C",
         "Information for internal use only, not for public release",
         "Minor operational impact, minor reputation damage",
         "Policies, procedures, org charts, internal communications",
         "3 years", "Biennial", "No"],
        ["L1", "PUBLIC", "🔵 PUBLIC", "#74C0FC",
         "Information approved for unrestricted distribution",
         "No impact expected",
         "Marketing materials, public website content, press releases",
         "As required", "Triennial", "No"],
    ]

    for row_idx, level in enumerate(levels, start=5):
        for col_idx, value in enumerate(level, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Apply level-specific styling to Level Name column
        level_name = level[1].lower()
        if level_name == "restricted":
            apply_style(ws.cell(row=row_idx, column=2), styles["level_restricted"])
        elif level_name == "confidential":
            apply_style(ws.cell(row=row_idx, column=2), styles["level_confidential"])
        elif level_name == "internal":
            apply_style(ws.cell(row=row_idx, column=2), styles["level_internal"])
        elif level_name == "public":
            apply_style(ws.cell(row=row_idx, column=2), styles["level_public"])

    # Add data validation for approval required
    dv_approval = DataValidation(
        type="list",
        formula1='"Yes - Executive,Yes - Manager,Yes - Owner,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_approval)
    dv_approval.add("J5:J20")

    set_column_widths(ws, {
        "A": 10, "B": 15, "C": 18, "D": 12,
        "E": 45, "F": 45, "G": 50,
        "H": 15, "I": 15, "J": 20
    })
    logger.info("Created Classification_Levels sheet")


def create_handling_requirements_sheet(ws, styles):
    """Create the Handling Requirements sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "Handling Requirements by Classification Level"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:L2")
    ws["A2"] = "Security controls and procedures required for each classification level"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Requirement Category", "RESTRICTED", "CONFIDENTIAL", "INTERNAL", "PUBLIC"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])
        if header == "RESTRICTED":
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        elif header == "CONFIDENTIAL":
            cell.fill = PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid")
        elif header == "INTERNAL":
            cell.fill = PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")
        elif header == "PUBLIC":
            cell.fill = PatternFill(start_color="74C0FC", end_color="74C0FC", fill_type="solid")

    requirements = [
        ["ACCESS CONTROL", "", "", "", ""],
        ["Need-to-know basis", "Mandatory", "Mandatory", "Recommended", "N/A"],
        ["Access approval required", "Executive + Owner", "Manager + Owner", "Owner", "None"],
        ["Access logging", "All access logged", "All access logged", "Modifications logged", "None required"],
        ["Access review frequency", "Quarterly", "Semi-annual", "Annual", "N/A"],
        ["", "", "", "", ""],
        ["STORAGE", "", "", "", ""],
        ["Encryption at rest", "AES-256 required", "AES-256 required", "Recommended", "Not required"],
        ["Storage location", "Approved systems only", "Approved systems only", "Corporate systems", "Any"],
        ["Personal device storage", "Prohibited", "Prohibited unless MDM", "With encryption", "Allowed"],
        ["Cloud storage", "Approved vendors only", "Approved vendors only", "Corporate cloud", "Any"],
        ["", "", "", "", ""],
        ["TRANSMISSION", "", "", "", ""],
        ["Encryption in transit", "TLS 1.3 required", "TLS 1.2+ required", "TLS recommended", "Not required"],
        ["Email transmission", "Encrypted + DLP", "Encrypted preferred", "Standard email", "Standard email"],
        ["External sharing", "Prohibited w/o approval", "Requires approval", "Discretion", "Allowed"],
        ["", "", "", "", ""],
        ["PHYSICAL HANDLING", "", "", "", ""],
        ["Printing", "Secure print only", "Track copies", "Normal", "Normal"],
        ["Clean desk", "Mandatory", "Mandatory", "Recommended", "N/A"],
        ["Secure disposal", "Shredding required", "Shredding required", "Shredding", "Recycling OK"],
        ["", "", "", "", ""],
        ["LABELLING", "", "", "", ""],
        ["Document marking", "Header/Footer + Watermark", "Header/Footer", "Footer only", "Optional"],
        ["Metadata tagging", "Mandatory", "Mandatory", "Recommended", "Optional"],
        ["Visual indicators", "Red banner", "Orange banner", "Green banner", "Blue banner"],
    ]

    for row_idx, req in enumerate(requirements, start=5):
        for col_idx, value in enumerate(req, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Bold category headers
            if col_idx == 1 and value.isupper() and value:
                cell.font = Font(bold=True)

    set_column_widths(ws, {"A": 25, "B": 25, "C": 25, "D": 25, "E": 25})
    logger.info("Created Handling_Requirements sheet")


def create_cia_matrix_sheet(ws, styles):
    """Create the CIA (Confidentiality, Integrity, Availability) Matrix sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "CIA Requirements Matrix"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "Define Confidentiality, Integrity, and Availability requirements per classification"
    apply_style(ws["A2"], styles["subheader"])

    # Confidentiality section
    ws.merge_cells("A4:G4")
    ws["A4"] = "CONFIDENTIALITY REQUIREMENTS"
    ws["A4"].font = Font(bold=True, size=12)

    c_headers = ["Level", "Access Control", "Encryption", "Disclosure Impact", "Monitoring"]
    for col, header in enumerate(c_headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["column_header"])

    c_data = [
        ["RESTRICTED", "Need-to-know, MFA required", "AES-256 at rest and transit",
         "Severe - regulatory/legal", "Real-time alerting"],
        ["CONFIDENTIAL", "Role-based, approval required", "AES-256 required",
         "Significant - business impact", "Daily review"],
        ["INTERNAL", "Department-based", "Recommended",
         "Minor - operational", "Weekly review"],
        ["PUBLIC", "Unrestricted", "Optional",
         "None expected", "None required"],
    ]

    for row_idx, data in enumerate(c_data, start=6):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True)

    # Integrity section
    ws.merge_cells("A12:G12")
    ws["A12"] = "INTEGRITY REQUIREMENTS"
    ws["A12"].font = Font(bold=True, size=12)

    i_headers = ["Level", "Change Control", "Version Control", "Modification Impact", "Validation"]
    for col, header in enumerate(i_headers, start=1):
        cell = ws.cell(row=13, column=col, value=header)
        apply_style(cell, styles["column_header"])

    i_data = [
        ["RESTRICTED", "Dual approval required", "Full audit trail",
         "Severe - may cause legal issues", "Digital signatures"],
        ["CONFIDENTIAL", "Manager approval", "Version history required",
         "Significant - incorrect decisions", "Checksums"],
        ["INTERNAL", "Owner approval", "Recommended",
         "Minor - process delays", "Spot checks"],
        ["PUBLIC", "Standard process", "Optional",
         "Minimal", "None required"],
    ]

    for row_idx, data in enumerate(i_data, start=14):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True)

    # Availability section
    ws.merge_cells("A20:G20")
    ws["A20"] = "AVAILABILITY REQUIREMENTS"
    ws["A20"].font = Font(bold=True, size=12)

    a_headers = ["Level", "Recovery Time", "Backup Frequency", "Unavailability Impact", "Redundancy"]
    for col, header in enumerate(a_headers, start=1):
        cell = ws.cell(row=21, column=col, value=header)
        apply_style(cell, styles["column_header"])

    a_data = [
        ["RESTRICTED", "< 4 hours", "Real-time/Hourly",
         "Critical - operations halt", "Active-active"],
        ["CONFIDENTIAL", "< 24 hours", "Daily",
         "Significant - major delays", "Active-passive"],
        ["INTERNAL", "< 72 hours", "Daily",
         "Minor - workarounds exist", "Standard backup"],
        ["PUBLIC", "Best effort", "Weekly",
         "Minimal", "Basic backup"],
    ]

    for row_idx, data in enumerate(a_data, start=22):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True)

    set_column_widths(ws, {"A": 15, "B": 25, "C": 25, "D": 25, "E": 20})
    logger.info("Created CIA_Matrix sheet")


def create_regulatory_mapping_sheet(ws, styles):
    """Create the Regulatory Mapping sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Classification to Regulatory Requirements Mapping"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = "Map classification levels to regulatory and legal requirements"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Regulation", "Requirement", "Data Types Covered",
        "Min Classification", "Special Handling", "Retention", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    mappings = [
        ["GDPR Art. 32", "Security of processing", "Personal data",
         "CONFIDENTIAL", "Pseudonymization where possible", "As per Art. 17", ""],
        ["GDPR Art. 9", "Special categories", "Health, biometric, genetic",
         "RESTRICTED", "Explicit consent required", "As per Art. 17", ""],
        ["Swiss nDSG Art. 8", "Data security", "Personal data (CH)",
         "CONFIDENTIAL", "Appropriate measures", "Purpose-based", ""],
        ["PCI DSS", "Cardholder data protection", "Credit card data",
         "RESTRICTED", "PCI controls required", "Per PCI standards", ""],
        ["SOX", "Financial integrity", "Financial records",
         "CONFIDENTIAL", "Audit trail required", "7 years", ""],
        ["HIPAA", "Health information", "PHI/ePHI",
         "RESTRICTED", "Minimum necessary", "6 years", ""],
        ["Trade Secrets", "Competitive advantage", "Proprietary information",
         "RESTRICTED", "NDA required", "Indefinite", ""],
        ["Employment Law", "Employee records", "HR data",
         "CONFIDENTIAL", "Access restricted to HR", "Per jurisdiction", ""],
    ]

    for row_idx, mapping in enumerate(mappings, start=5):
        for col_idx, value in enumerate(mapping, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True)
            if col_idx == 7:  # Status column
                cell.fill = styles["input_cell"]["fill"]

    dv_status = DataValidation(
        type="list",
        formula1='"✅ Compliant,⚠️ Partial,❌ Non-Compliant,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("G5:G30")

    dv_level = DataValidation(
        type="list",
        formula1='"RESTRICTED,CONFIDENTIAL,INTERNAL,PUBLIC"',
        allow_blank=True
    )
    ws.add_data_validation(dv_level)
    dv_level.add("D5:D30")

    set_column_widths(ws, {
        "A": 18, "B": 25, "C": 25, "D": 18, "E": 30, "F": 18, "G": 18
    })
    logger.info("Created Regulatory_Mapping sheet")


def create_evidence_register_sheet(ws, styles):
    """Create the Evidence Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Evidence Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Track audit evidence supporting classification scheme implementation"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Evidence ID", "Description", "Evidence Type",
        "Related Requirement", "Location", "Collected Date",
        "Collected By", "Verification Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty rows with styling
    for row in range(5, 25):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            if col in [1, 3, 7, 8]:
                cell.fill = styles["input_cell"]["fill"]

    dv_type = DataValidation(
        type="list",
        formula1='"Policy document,Procedure document,Configuration export,Screenshot,Training record,Acknowledgment,Audit report,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C5:C30")

    dv_status = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending Review,❌ Not Verified,Expired"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H30")

    set_column_widths(ws, {
        "A": 15, "B": 40, "C": 20, "D": 25, "E": 30, "F": 15, "G": 15, "H": 18
    })
    logger.info("Created Evidence_Register sheet")


def create_approval_sheet(ws, styles):
    """Create the Approval Sign-Off sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "Classification Scheme Approval"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "Formal approval record for classification scheme"
    apply_style(ws["A2"], styles["subheader"])

    # Scheme metadata
    ws["A4"] = "Scheme Version:"
    ws["B4"] = "1.0"
    ws["A5"] = "Effective Date:"
    ws["B5"] = ""
    ws["A6"] = "Next Review Date:"
    ws["B6"] = ""
    ws["A7"] = "Scheme Owner:"
    ws["B7"] = ""

    for row in range(4, 8):
        ws[f"A{row}"].font = Font(bold=True)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]

    # Approval table
    ws["A10"] = "APPROVAL SIGNATURES"
    ws["A10"].font = Font(bold=True, size=12)

    approval_headers = ["Role", "Name", "Signature", "Date", "Decision", "Comments"]
    for col, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=12, column=col, value=header)
        apply_style(cell, styles["column_header"])

    approvers = [
        "Chief Information Security Officer",
        "Data Protection Officer",
        "Chief Information Officer",
        "Legal/Compliance Representative",
        "Business Unit Representative",
    ]

    for row_idx, role in enumerate(approvers, start=13):
        ws.cell(row=row_idx, column=1, value=role).border = styles["border"]
        for col in range(2, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected,Pending"',
        allow_blank=True
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add("E13:E20")

    set_column_widths(ws, {"A": 35, "B": 25, "C": 20, "D": 15, "E": 22, "F": 30})
    logger.info("Created Approval_SignOff sheet")


# =============================================================================
# MAIN FUNCTION
# =============================================================================
def main():
    """Main entry point."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS-IMP-A.5.12-13.S1 Classification Scheme Definition Generator")
        logger.info("=" * 70)

        wb = Workbook()
        styles = setup_styles()

        # Rename default sheet
        ws_instructions = wb.active
        ws_instructions.title = "Instructions"

        # Create all sheets
        ws_levels = wb.create_sheet("Classification_Levels")
        ws_handling = wb.create_sheet("Handling_Requirements")
        ws_cia = wb.create_sheet("CIA_Matrix")
        ws_regulatory = wb.create_sheet("Regulatory_Mapping")
        ws_evidence = wb.create_sheet("Evidence_Register")
        ws_approval = wb.create_sheet("Approval_SignOff")

        # Populate sheets
        create_instructions_sheet(ws_instructions, styles)
        create_classification_levels_sheet(ws_levels, styles)
        create_handling_requirements_sheet(ws_handling, styles)
        create_cia_matrix_sheet(ws_cia, styles)
        create_regulatory_mapping_sheet(ws_regulatory, styles)
        create_evidence_register_sheet(ws_evidence, styles)
        create_approval_sheet(ws_approval, styles)

        # Save workbook
        wb.save(OUTPUT_FILENAME)
        logger.info("=" * 70)
        logger.info("SUCCESS: Workbook saved as %s", OUTPUT_FILENAME)
        logger.info("=" * 70)
        return 0

    except Exception as e:
        logger.error("FAILED: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.12-13 Classification & Labelling
# =============================================================================
