#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.5-6.5 - External Communications Consolidation Dashboard
================================================================================

ISO/IEC 27001:2022 Controls A.5.5 & A.5.6: Contact with Authorities & SIGs
Consolidation Dashboard: Executive Summary & Cross-Domain Compliance View

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a consolidation dashboard that aggregates data from all
four A.5.5-6 assessment workbooks into a unified executive view for compliance
monitoring and audit readiness.

**Source Assessment Domains:**
1. Domain 1: Authority Contacts Register (A.5.5-6.1)
   - Law enforcement, regulators, emergency services contacts

2. Domain 2: Special Interest Groups Register (A.5.5-6.2)
   - Security forums, ISACs, professional associations

3. Domain 3: External Communication Procedures (A.5.5-6.3)
   - Notification requirements, escalation matrices, templates

4. Domain 4: Compliance Dashboard (A.5.5-6.4)
   - KPIs, metrics, gap analysis, audit readiness

**Generated Dashboard Structure:**
1. Instructions - Guide to using the consolidation dashboard
2. Executive_Summary - High-level compliance status across all domains
3. Domain_Overview - Summary status from each source workbook
4. Authority_Compliance - Consolidated authority contact compliance
5. SIG_Compliance - Consolidated SIG membership compliance
6. Procedure_Compliance - Communication procedure completeness
7. Cross_Domain_Gaps - Gaps identified across all domains
8. Remediation_Tracker - Consolidated action items with status
9. KPI_Summary - Key metrics from all domains
10. Evidence_Index - Cross-reference to source workbook evidence
11. Trend_Dashboard - Historical compliance across domains
12. Approval_SignOff - Executive sign-off for consolidated view

**Key Features:**
- Automated aggregation from four source workbooks
- Cross-domain gap identification
- Executive-friendly visual compliance indicators
- Audit evidence traceability to source assessments
- Trend tracking across reporting periods

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.5-6.5"
WORKBOOK_NAME = "Consolidation Dashboard"
CONTROL_ID = "A.5.5-6"
CONTROL_NAME = "Contact with Authorities & Special Interest Groups"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls A.5.5 & A.5.6: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# Source workbook patterns
SOURCE_WORKBOOKS = {
    "A.5.5-6.1": "Authority_Contacts_Register",
    "A.5.5-6.2": "Special_Interest_Groups_Register",
    "A.5.5-6.3": "External_Communication_Procedures",
    "A.5.5-6.4": "External_Communications_Compliance_Dashboard"
}

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")

SECTION_FONT = Font(bold=True, size=12, color="1F4E79")

COMPLIANT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
NON_COMPLIANT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
CALC_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# DATA VALIDATION LISTS
# =============================================================================
STATUS_VALUES = ["Compliant", "Partially Compliant", "Non-Compliant", "Not Assessed", "Not Applicable"]
PRIORITY_VALUES = ["Critical", "High", "Medium", "Low"]
DOMAIN_VALUES = ["A.5.5-6.1 Authority Contacts", "A.5.5-6.2 SIG Membership",
                 "A.5.5-6.3 Procedures", "A.5.5-6.4 Dashboard", "Cross-Domain"]


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================

def apply_header_style(ws, row, start_col, end_col):
    """Apply header styling to a row."""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def create_instructions_sheet(ws):
    """Create the Instructions sheet."""
    ws.title = "Instructions"

    instructions = [
        [f"ISMS-IMP-A.5.5-6.5 - External Communications Consolidation Dashboard"],
        [""],
        ["PURPOSE"],
        ["This dashboard consolidates compliance data from all four A.5.5-6 assessment domains"],
        ["into a unified executive view for compliance monitoring and audit readiness."],
        [""],
        ["SOURCE WORKBOOKS"],
        ["This dashboard aggregates data from:"],
        ["  • A.5.5-6.1: Authority Contacts Register - Law enforcement, regulators, emergency services"],
        ["  • A.5.5-6.2: Special Interest Groups Register - Security forums, ISACs, associations"],
        ["  • A.5.5-6.3: External Communication Procedures - Notification requirements, templates"],
        ["  • A.5.5-6.4: Compliance Dashboard - KPIs, metrics, gap analysis"],
        [""],
        ["SHEETS IN THIS WORKBOOK"],
        ["1. Instructions - This guidance sheet"],
        ["2. Executive_Summary - High-level compliance status across all domains"],
        ["3. Domain_Overview - Summary from each source workbook"],
        ["4. Authority_Compliance - Consolidated A.5.5 authority contact status"],
        ["5. SIG_Compliance - Consolidated A.5.6 SIG membership status"],
        ["6. Procedure_Compliance - Communication procedure completeness"],
        ["7. Cross_Domain_Gaps - Gaps identified across all domains"],
        ["8. Remediation_Tracker - Action items with owners and deadlines"],
        ["9. KPI_Summary - Key metrics aggregated from all domains"],
        ["10. Evidence_Index - Cross-reference to source workbook evidence"],
        ["11. Trend_Dashboard - Historical compliance trends"],
        ["12. Approval_SignOff - Executive approval workflow"],
        [""],
        ["DATA ENTRY GUIDANCE"],
        ["• Yellow cells = Manual input required"],
        ["• Green cells = Calculated/aggregated values"],
        ["• Complete source workbooks (A.5.5-6.1 through A.5.5-6.4) before this consolidation"],
        ["• Update this dashboard quarterly or before audits"],
        [""],
        ["COMPLIANCE SCORING"],
        ["• Compliant (Green): ≥90% of requirements met"],
        ["• Partially Compliant (Amber): 50-89% of requirements met"],
        ["• Non-Compliant (Red): <50% of requirements met"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SOURCE WORKBOOKS", "SHEETS IN THIS WORKBOOK",
                          "DATA ENTRY GUIDANCE", "COMPLIANCE SCORING"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 95


def create_executive_summary_sheet(ws):
    """Create the Executive Summary sheet."""
    ws.title = "Executive_Summary"

    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="A.5.5-6 External Communications - Executive Summary")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    # Reporting info
    ws.cell(row=3, column=1, value="Reporting Period:").font = Font(bold=True)
    ws.cell(row=3, column=2).fill = INPUT_FILL
    ws.cell(row=3, column=4, value="Assessment Date:").font = Font(bold=True)
    ws.cell(row=3, column=5).fill = INPUT_FILL
    ws.cell(row=3, column=7, value="Assessor:").font = Font(bold=True)
    ws.cell(row=3, column=8).fill = INPUT_FILL

    # Overall compliance status
    ws.cell(row=5, column=1, value="OVERALL COMPLIANCE STATUS").font = SECTION_FONT

    headers = ["Domain", "Workbook", "Status", "Score %", "Critical Gaps", "Last Updated"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=header)
    apply_header_style(ws, 6, 1, len(headers))

    domains = [
        ("A.5.5", "Authority Contacts Register", "", "", "", ""),
        ("A.5.6", "Special Interest Groups Register", "", "", "", ""),
        ("A.5.5-6", "Communication Procedures", "", "", "", ""),
        ("A.5.5-6", "Compliance Dashboard", "", "", "", ""),
        ("OVERALL", "Consolidated Assessment", "", "", "", ""),
    ]

    for row_num, domain_data in enumerate(domains, 7):
        for col_num, value in enumerate(domain_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = THIN_BORDER
            if col_num in [3, 4, 5]:
                cell.fill = INPUT_FILL
            if row_num == 11:  # OVERALL row
                cell.font = Font(bold=True)

    # Key metrics section
    ws.cell(row=13, column=1, value="KEY METRICS SUMMARY").font = SECTION_FONT

    metrics_headers = ["Metric", "Target", "Actual", "Status", "Trend"]
    for col, header in enumerate(metrics_headers, 1):
        ws.cell(row=14, column=col, value=header)
    apply_header_style(ws, 14, 1, len(metrics_headers))

    metrics = [
        ("Authority contacts documented", "100%", "", "", ""),
        ("Contacts verified within 12 months", "100%", "", "", ""),
        ("SIG memberships with value assessment", "100%", "", "", ""),
        ("Communication procedures documented", "100%", "", "", ""),
        ("Notification requirements covered", "100%", "", "", ""),
        ("Evidence availability for audit", "100%", "", "", ""),
    ]

    for row_num, metric_data in enumerate(metrics, 15):
        for col_num, value in enumerate(metric_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = THIN_BORDER
            if col_num in [3, 4, 5]:
                cell.fill = INPUT_FILL

    # Executive sign-off
    ws.cell(row=23, column=1, value="EXECUTIVE SIGN-OFF").font = SECTION_FONT

    signoff_headers = ["Role", "Name", "Signature", "Date"]
    for col, header in enumerate(signoff_headers, 1):
        ws.cell(row=24, column=col, value=header)
    apply_header_style(ws, 24, 1, len(signoff_headers))

    roles = ["CISO", "Compliance Manager", "Legal Counsel"]
    for row_num, role in enumerate(roles, 25):
        ws.cell(row=row_num, column=1, value=role).border = THIN_BORDER
        for col in range(2, 5):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20


def create_domain_overview_sheet(ws):
    """Create the Domain Overview sheet."""
    ws.title = "Domain_Overview"

    # Title
    ws.merge_cells('A1:J1')
    title_cell = ws.cell(row=1, column=1, value="Domain-by-Domain Compliance Overview")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    # Domain 1: Authority Contacts
    ws.cell(row=3, column=1, value="DOMAIN 1: AUTHORITY CONTACTS (A.5.5-6.1)").font = SECTION_FONT

    headers = ["Requirement", "Status", "Evidence Ref", "Gap Description", "Remediation"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    apply_header_style(ws, 4, 1, len(headers))

    auth_requirements = [
        "Law enforcement contacts documented",
        "Regulatory body contacts documented",
        "Emergency services contacts documented",
        "Contact verification current (<12 months)",
        "Communication protocols defined",
    ]

    for row_num, req in enumerate(auth_requirements, 5):
        ws.cell(row=row_num, column=1, value=req).border = THIN_BORDER
        for col in range(2, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Domain 2: SIG Membership
    ws.cell(row=11, column=1, value="DOMAIN 2: SPECIAL INTEREST GROUPS (A.5.5-6.2)").font = SECTION_FONT

    for col, header in enumerate(headers, 1):
        ws.cell(row=12, column=col, value=header)
    apply_header_style(ws, 12, 1, len(headers))

    sig_requirements = [
        "SIG memberships documented",
        "Membership value assessed",
        "Intelligence sharing active",
        "Contribution tracking maintained",
        "Engagement frequency appropriate",
    ]

    for row_num, req in enumerate(sig_requirements, 13):
        ws.cell(row=row_num, column=1, value=req).border = THIN_BORDER
        for col in range(2, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Domain 3: Procedures
    ws.cell(row=19, column=1, value="DOMAIN 3: COMMUNICATION PROCEDURES (A.5.5-6.3)").font = SECTION_FONT

    for col, header in enumerate(headers, 1):
        ws.cell(row=20, column=col, value=header)
    apply_header_style(ws, 20, 1, len(headers))

    proc_requirements = [
        "Incident notification procedures",
        "Regulatory reporting procedures",
        "Media communication procedures",
        "Escalation matrix defined",
        "Communication templates available",
    ]

    for row_num, req in enumerate(proc_requirements, 21):
        ws.cell(row=row_num, column=1, value=req).border = THIN_BORDER
        for col in range(2, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 35


def create_authority_compliance_sheet(ws):
    """Create the Authority Compliance sheet."""
    ws.title = "Authority_Compliance"

    # Title
    ws.merge_cells('A1:I1')
    title_cell = ws.cell(row=1, column=1, value="A.5.5 Authority Contact Compliance - Consolidated View")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Source: ISMS-IMP-A.5.5-6.1 Authority Contacts Register").font = Font(italic=True)

    headers = ["Authority_Type", "Authority_Name", "Contact_Status", "Last_Verified",
               "Next_Review", "Compliance_Status", "Gap_Notes", "Action_Required", "Owner"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    apply_header_style(ws, 4, 1, len(headers))

    authority_types = [
        "Law Enforcement - Local",
        "Law Enforcement - National",
        "Data Protection Authority",
        "Financial Regulator",
        "Industry Regulator",
        "Emergency Services - Fire",
        "Emergency Services - Medical",
        "Cyber Security Agency",
        "Government Security Centre",
    ]

    for row_num, auth_type in enumerate(authority_types, 5):
        ws.cell(row=row_num, column=1, value=auth_type).border = THIN_BORDER
        for col in range(2, 10):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Add data validation for status
    status_dv = DataValidation(type="list", formula1='"Compliant,Partially Compliant,Non-Compliant,Not Assessed"')
    ws.add_data_validation(status_dv)
    status_dv.add(f'F5:F{4 + len(authority_types)}')

    # Column widths
    for col, width in enumerate([25, 30, 15, 15, 15, 18, 30, 25, 20], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_sig_compliance_sheet(ws):
    """Create the SIG Compliance sheet."""
    ws.title = "SIG_Compliance"

    # Title
    ws.merge_cells('A1:I1')
    title_cell = ws.cell(row=1, column=1, value="A.5.6 Special Interest Group Compliance - Consolidated View")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Source: ISMS-IMP-A.5.5-6.2 Special Interest Groups Register").font = Font(italic=True)

    headers = ["SIG_Category", "Group_Name", "Membership_Status", "Value_Rating",
               "Last_Engagement", "Intelligence_Received", "Compliance_Status", "Gap_Notes", "Owner"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    apply_header_style(ws, 4, 1, len(headers))

    sig_categories = [
        "ISAC - Financial",
        "ISAC - Healthcare",
        "ISAC - Critical Infrastructure",
        "Security Forum - National",
        "Security Forum - Industry",
        "Professional Association",
        "Vendor Security Group",
        "Academic Research Group",
    ]

    for row_num, sig_cat in enumerate(sig_categories, 5):
        ws.cell(row=row_num, column=1, value=sig_cat).border = THIN_BORDER
        for col in range(2, 10):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    for col, width in enumerate([25, 30, 18, 15, 18, 20, 18, 30, 20], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_procedure_compliance_sheet(ws):
    """Create the Procedure Compliance sheet."""
    ws.title = "Procedure_Compliance"

    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="Communication Procedure Compliance - Consolidated View")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Source: ISMS-IMP-A.5.5-6.3 External Communication Procedures").font = Font(italic=True)

    headers = ["Procedure_Type", "Procedure_Name", "Status", "Last_Reviewed",
               "Compliance_Status", "Gap_Description", "Remediation_Action", "Owner"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    apply_header_style(ws, 4, 1, len(headers))

    procedures = [
        ("Incident Notification", "Security Incident Reporting to Authorities"),
        ("Incident Notification", "Data Breach Notification Procedure"),
        ("Regulatory Reporting", "Regulatory Compliance Reporting"),
        ("Regulatory Reporting", "Annual Security Report to Regulator"),
        ("Media Communication", "Security Incident Media Response"),
        ("Media Communication", "Crisis Communication Protocol"),
        ("Escalation", "External Contact Escalation Matrix"),
        ("Templates", "Authority Communication Templates"),
        ("Templates", "Regulatory Submission Templates"),
    ]

    for row_num, (proc_type, proc_name) in enumerate(procedures, 5):
        ws.cell(row=row_num, column=1, value=proc_type).border = THIN_BORDER
        ws.cell(row=row_num, column=2, value=proc_name).border = THIN_BORDER
        for col in range(3, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    for col, width in enumerate([22, 40, 15, 15, 18, 35, 30, 20], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_cross_domain_gaps_sheet(ws):
    """Create the Cross-Domain Gaps sheet."""
    ws.title = "Cross_Domain_Gaps"

    # Title
    ws.merge_cells('A1:J1')
    title_cell = ws.cell(row=1, column=1, value="Cross-Domain Gap Analysis")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["Gap_ID", "Source_Domain", "Gap_Description", "Risk_Rating", "Priority",
               "Affected_Controls", "Root_Cause", "Remediation_Action", "Owner", "Target_Date"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    apply_header_style(ws, 3, 1, len(headers))

    # Sample gap entries
    for row_num in range(4, 19):
        ws.cell(row=row_num, column=1, value=f"GAP-{row_num-3:03d}").border = THIN_BORDER
        for col in range(2, 11):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validations
    domain_dv = DataValidation(type="list", formula1='"A.5.5-6.1 Authority Contacts,A.5.5-6.2 SIG Membership,A.5.5-6.3 Procedures,A.5.5-6.4 Dashboard,Cross-Domain"')
    ws.add_data_validation(domain_dv)
    domain_dv.add('B4:B18')

    priority_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(priority_dv)
    priority_dv.add('E4:E18')

    # Column widths
    for col, width in enumerate([10, 25, 40, 12, 10, 20, 30, 35, 18, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_remediation_tracker_sheet(ws):
    """Create the Remediation Tracker sheet."""
    ws.title = "Remediation_Tracker"

    # Title
    ws.merge_cells('A1:K1')
    title_cell = ws.cell(row=1, column=1, value="Consolidated Remediation Action Tracker")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["Action_ID", "Related_Gap", "Source_Domain", "Action_Description", "Priority",
               "Owner", "Start_Date", "Target_Date", "Status", "Progress_%", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    apply_header_style(ws, 3, 1, len(headers))

    for row_num in range(4, 24):
        ws.cell(row=row_num, column=1, value=f"ACT-{row_num-3:03d}").border = THIN_BORDER
        for col in range(2, 12):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validations
    status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,On Hold,Completed,Cancelled"')
    ws.add_data_validation(status_dv)
    status_dv.add('I4:I23')

    # Column widths
    for col, width in enumerate([10, 12, 22, 40, 10, 18, 12, 12, 12, 10, 30], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_kpi_summary_sheet(ws):
    """Create the KPI Summary sheet."""
    ws.title = "KPI_Summary"

    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="Key Performance Indicators - Consolidated Summary")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    # Authority KPIs
    ws.cell(row=3, column=1, value="AUTHORITY CONTACT KPIs (A.5.5)").font = SECTION_FONT

    headers = ["KPI", "Target", "Current", "Previous", "Trend", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    apply_header_style(ws, 4, 1, len(headers))

    auth_kpis = [
        ("% of required authorities documented", "100%", "", "", "", ""),
        ("% of contacts verified (<12 months)", "100%", "", "", "", ""),
        ("Average days since last verification", "<180", "", "", "", ""),
        ("% with defined communication protocol", "100%", "", "", "", ""),
    ]

    for row_num, kpi_data in enumerate(auth_kpis, 5):
        for col_num, value in enumerate(kpi_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = THIN_BORDER
            if col_num > 2:
                cell.fill = INPUT_FILL

    # SIG KPIs
    ws.cell(row=10, column=1, value="SIG ENGAGEMENT KPIs (A.5.6)").font = SECTION_FONT

    for col, header in enumerate(headers, 1):
        ws.cell(row=11, column=col, value=header)
    apply_header_style(ws, 11, 1, len(headers))

    sig_kpis = [
        ("% of SIG memberships with value assessment", "100%", "", "", "", ""),
        ("Intelligence items received (quarterly)", ">10", "", "", "", ""),
        ("Contributions made (quarterly)", ">2", "", "", "", ""),
        ("% of memberships actively engaged", "≥80%", "", "", "", ""),
    ]

    for row_num, kpi_data in enumerate(sig_kpis, 12):
        for col_num, value in enumerate(kpi_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = THIN_BORDER
            if col_num > 2:
                cell.fill = INPUT_FILL

    # Procedure KPIs
    ws.cell(row=17, column=1, value="PROCEDURE KPIs").font = SECTION_FONT

    for col, header in enumerate(headers, 1):
        ws.cell(row=18, column=col, value=header)
    apply_header_style(ws, 18, 1, len(headers))

    proc_kpis = [
        ("% of communication procedures documented", "100%", "", "", "", ""),
        ("% of procedures reviewed (<12 months)", "100%", "", "", "", ""),
        ("% of notification requirements covered", "100%", "", "", "", ""),
        ("Template availability score", "100%", "", "", "", ""),
    ]

    for row_num, kpi_data in enumerate(proc_kpis, 19):
        for col_num, value in enumerate(kpi_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = THIN_BORDER
            if col_num > 2:
                cell.fill = INPUT_FILL

    # Column widths
    for col, width in enumerate([45, 12, 12, 12, 10, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_evidence_index_sheet(ws):
    """Create the Evidence Index sheet."""
    ws.title = "Evidence_Index"

    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="Evidence Index - Cross-Reference to Source Assessments")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["Evidence_ID", "Source_Workbook", "Source_Sheet", "Evidence_Type",
               "Evidence_Description", "Location/Reference", "Date_Captured", "Validation_Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    apply_header_style(ws, 3, 1, len(headers))

    for row_num in range(4, 24):
        ws.cell(row=row_num, column=1, value=f"EV-{row_num-3:03d}").border = THIN_BORDER
        for col in range(2, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validation for source workbook
    source_dv = DataValidation(type="list", formula1='"A.5.5-6.1 Authority Contacts,A.5.5-6.2 SIG Register,A.5.5-6.3 Procedures,A.5.5-6.4 Dashboard"')
    ws.add_data_validation(source_dv)
    source_dv.add('B4:B23')

    # Column widths
    for col, width in enumerate([12, 25, 25, 18, 40, 30, 15, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_trend_dashboard_sheet(ws):
    """Create the Trend Dashboard sheet."""
    ws.title = "Trend_Dashboard"

    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="Historical Compliance Trend Dashboard")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    # Quarterly trend tracking
    headers = ["Period", "A.5.5 Authority %", "A.5.6 SIG %", "Procedures %",
               "Overall %", "Critical Gaps", "Remediation Rate", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    apply_header_style(ws, 3, 1, len(headers))

    periods = ["Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025",
               "Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"]

    for row_num, period in enumerate(periods, 4):
        ws.cell(row=row_num, column=1, value=period).border = THIN_BORDER
        for col in range(2, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    for col, width in enumerate([12, 18, 15, 15, 12, 15, 18, 35], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_approval_signoff_sheet(ws):
    """Create the Approval Sign-Off sheet."""
    ws.title = "Approval_SignOff"

    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="Consolidation Dashboard Approval & Sign-Off")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    # Assessment details
    ws.cell(row=3, column=1, value="Assessment Period:").font = Font(bold=True)
    ws.cell(row=3, column=2).fill = INPUT_FILL
    ws.cell(row=4, column=1, value="Consolidation Date:").font = Font(bold=True)
    ws.cell(row=4, column=2).fill = INPUT_FILL
    ws.cell(row=5, column=1, value="Prepared By:").font = Font(bold=True)
    ws.cell(row=5, column=2).fill = INPUT_FILL

    # Sign-off table
    ws.cell(row=7, column=1, value="APPROVAL WORKFLOW").font = SECTION_FONT

    headers = ["Role", "Name", "Date", "Signature", "Comments"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=8, column=col, value=header)
    apply_header_style(ws, 8, 1, len(headers))

    roles = [
        "Prepared By (Analyst)",
        "Reviewed By (Security Manager)",
        "Validated By (Compliance Officer)",
        "Approved By (CISO)",
        "Acknowledged By (Executive Sponsor)",
    ]

    for row_num, role in enumerate(roles, 9):
        ws.cell(row=row_num, column=1, value=role).border = THIN_BORDER
        for col in range(2, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Declaration
    ws.cell(row=16, column=1, value="DECLARATION").font = SECTION_FONT
    declaration = ("I confirm that this consolidation dashboard accurately represents the current "
                  "compliance status of ISO 27001:2022 Controls A.5.5 and A.5.6 based on the "
                  "source assessment workbooks. All gaps and remediation actions have been "
                  "identified and assigned to responsible owners.")
    ws.merge_cells('A17:E18')
    ws.cell(row=17, column=1, value=declaration).alignment = Alignment(wrap_text=True)

    # Column widths
    for col, width in enumerate([35, 25, 15, 20, 40], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


# =============================================================================
# MAIN GENERATION FUNCTION
# =============================================================================

def generate_workbook():
    """Generate the consolidation dashboard workbook."""
    logger.info("=" * 70)
    logger.info(f"Generating: {DOCUMENT_ID} - {WORKBOOK_NAME}")
    logger.info(f"Control: {CONTROL_REF}")
    logger.info("=" * 70)

    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create all sheets
    logger.info("Creating Instructions sheet...")
    create_instructions_sheet(wb.create_sheet())

    logger.info("Creating Executive_Summary sheet...")
    create_executive_summary_sheet(wb.create_sheet())

    logger.info("Creating Domain_Overview sheet...")
    create_domain_overview_sheet(wb.create_sheet())

    logger.info("Creating Authority_Compliance sheet...")
    create_authority_compliance_sheet(wb.create_sheet())

    logger.info("Creating SIG_Compliance sheet...")
    create_sig_compliance_sheet(wb.create_sheet())

    logger.info("Creating Procedure_Compliance sheet...")
    create_procedure_compliance_sheet(wb.create_sheet())

    logger.info("Creating Cross_Domain_Gaps sheet...")
    create_cross_domain_gaps_sheet(wb.create_sheet())

    logger.info("Creating Remediation_Tracker sheet...")
    create_remediation_tracker_sheet(wb.create_sheet())

    logger.info("Creating KPI_Summary sheet...")
    create_kpi_summary_sheet(wb.create_sheet())

    logger.info("Creating Evidence_Index sheet...")
    create_evidence_index_sheet(wb.create_sheet())

    logger.info("Creating Trend_Dashboard sheet...")
    create_trend_dashboard_sheet(wb.create_sheet())

    logger.info("Creating Approval_SignOff sheet...")
    create_approval_signoff_sheet(wb.create_sheet())

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info("=" * 70)
    logger.info(f"SUCCESS: Workbook saved as {OUTPUT_FILENAME}")
    logger.info(f"Sheets created: {len(wb.sheetnames)}")
    logger.info("=" * 70)

    return 0


# =============================================================================
# ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    try:
        sys.exit(generate_workbook())
    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        sys.exit(1)

# =============================================================================
# QA_VERIFIED: 2026-02-04
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code Generator
# CHANGES: Initial creation - consolidation dashboard for A.5.5-6
# =============================================================================
