#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
A.5.17 Workbook Normalizer - Normalizes all A.5.17 Authentication Information
workbooks for consistent formatting.

Usage: python3 normalize_a517.py
"""

import logging
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

CONTROL_ID = "A.5.17"
WORKBOOK_DIR = Path(__file__).parent.parent / "90_workbooks"
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")


def normalize_workbook(filepath: Path) -> bool:
    try:
        logger.info(f"Normalizing: {filepath.name}")
        wb = load_workbook(filepath)
        for ws in wb.worksheets:
            if ws["A1"].value:
                ws["A1"].font = HEADER_FONT
                ws["A1"].fill = HEADER_FILL
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if ws.freeze_panes is None and ws.max_row > 3:
                ws.freeze_panes = "A4"
        wb.save(filepath)
        logger.info(f"  ✓ Normalized successfully")
        return True
    except Exception as e:
        logger.error(f"  ✗ Error: {e}")
        return False


def main() -> int:
    logger.info("=" * 70)
    logger.info(f"{CONTROL_ID} Workbook Normalizer")
    logger.info("=" * 70)

    if not WORKBOOK_DIR.exists():
        logger.warning(f"Workbook directory not found: {WORKBOOK_DIR}")
        return 0

    workbooks = list(WORKBOOK_DIR.glob("ISMS-IMP-A.5.17*.xlsx"))
    if not workbooks:
        logger.info("No A.5.17 workbooks found")
        return 0

    logger.info(f"Found {len(workbooks)} workbook(s)")
    success = sum(1 for wb in workbooks if normalize_workbook(wb))
    failed = len(workbooks) - success

    logger.info("=" * 70)
    logger.info(f"Results: {success} normalized, {failed} failed")
    logger.info("=" * 70)
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.17 Authentication Information control
# =============================================================================
