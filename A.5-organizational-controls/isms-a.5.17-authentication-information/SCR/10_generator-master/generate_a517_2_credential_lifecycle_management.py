#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
A.5.17.2 Credential Lifecycle Management Generator
================================================================================

Generates Excel workbook for managing the complete lifecycle of authentication
credentials including allocation, change, recovery, and revocation per ISO 27001:2022.

Sheets:
    1. Instructions - Completion guidance
    2. Allocation_Process - New credential issuance procedures
    3. Change_Management - Password/credential change procedures
    4. Recovery_Process - Self-service and assisted recovery
    5. Revocation_Process - Credential revocation and termination
    6. Audit_Log_Requirements - Authentication event logging
    7. Evidence_Register - Supporting documentation
    8. Approval_SignOff - Authorization workflow

Usage:
    python3 generate_a517_2_credential_lifecycle_management.py
================================================================================
"""

import logging
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

DOCUMENT_ID = "ISMS-IMP-A.5.17.2"
WORKBOOK_NAME = "Credential Lifecycle Management"
CONTROL_ID = "A.5.17"
CONTROL_NAME = "Authentication Information"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP_LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


def set_column_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_header_row(ws, row: int, headers: list):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def add_data_validation(ws, cell_range: str, formula: str):
    dv = DataValidation(type="list", formula1=formula, showDropDown=False, allowBlank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def create_instructions_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Instructions"

    ws.merge_cells("A1:H1")
    ws["A1"].value = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    metadata = [
        ("Document ID:", DOCUMENT_ID),
        ("Control Reference:", CONTROL_REF),
        ("Generated Date:", GENERATED_DATE),
        ("Version:", "1.0"),
    ]

    row = 3
    for label, value in metadata:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="PURPOSE").font = BOLD_FONT
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1).value = (
        "This workbook documents credential lifecycle management processes per ISO 27001:2022 A.5.17. "
        "It covers credential allocation, change management, recovery procedures, and revocation processes. "
        "Use this to ensure consistent and secure handling of authentication credentials throughout their lifecycle."
    )
    ws.cell(row=row, column=1).font = NORMAL_FONT
    ws.cell(row=row, column=1).alignment = TOP_LEFT_ALIGN
    ws.row_dimensions[row].height = 50

    row += 2
    ws.cell(row=row, column=1, value="SHEET DESCRIPTIONS").font = BOLD_FONT
    row += 1

    sheets = [
        ("Allocation_Process", "New user credential issuance and enrollment"),
        ("Change_Management", "Password/credential change procedures"),
        ("Recovery_Process", "Self-service and assisted recovery workflows"),
        ("Revocation_Process", "Credential revocation and account termination"),
        ("Audit_Log_Requirements", "Authentication event logging requirements"),
        ("Evidence_Register", "Supporting documentation"),
        ("Approval_SignOff", "Process approval workflow"),
    ]

    for sheet_name, description in sheets:
        ws.cell(row=row, column=1, value=sheet_name).font = BOLD_FONT
        ws.cell(row=row, column=2, value=description).font = NORMAL_FONT
        row += 1

    set_column_widths(ws, {"A": 25, "B": 55, "C": 20, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20})
    ws.freeze_panes = "A3"
    logger.info("Created Instructions sheet")


def create_allocation_process_sheet(wb: Workbook):
    ws = wb.create_sheet("Allocation_Process")

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Credential Allocation Process"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["Process Step", "Description", "Responsible Role", "Verification Required", "SLA", "System/Tool", "Status", "Notes"]
    create_header_row(ws, 3, headers)

    steps = [
        ("1. Access Request", "User/manager submits access request via approved channel", "Requestor", "Manager approval", "Same day", "ServiceNow/ITSM", "", ""),
        ("2. Identity Verification", "Verify user identity before credential issuance", "IT/HR", "Photo ID, employment verification", "Same day", "HR System", "", ""),
        ("3. Authorization Check", "Verify user is authorized for requested access", "System Owner", "Role-based access approval", "1 business day", "IAM System", "", ""),
        ("4. Account Creation", "Create user account in target system", "IT Operations", "Naming convention compliance", "Same day", "Active Directory", "", ""),
        ("5. Initial Password", "Generate temporary password meeting policy", "Automated/IT", "Complexity requirements", "Immediate", "Password Generator", "", ""),
        ("6. Secure Delivery", "Transmit initial credentials via secure channel", "IT Operations", "Encrypted transmission", "Same day", "Secure Email/Portal", "", ""),
        ("7. MFA Enrollment", "Enroll user in required MFA methods", "User/IT Support", "MFA device registered", "Within 24 hours", "MFA System", "", ""),
        ("8. First Login", "User changes temporary password on first login", "User", "Password changed successfully", "Within 24 hours", "Target System", "", ""),
        ("9. Confirmation", "Confirm successful access and close request", "Requestor", "User confirms access", "Within 48 hours", "ServiceNow/ITSM", "", ""),
        ("10. Documentation", "Record credential allocation event", "System", "Audit log entry", "Immediate", "SIEM/Audit Log", "", ""),
    ]

    row = 4
    for step in steps:
        for col, value in enumerate(step, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [7, 8]:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"G4:G{row+5}", '"Implemented,In Progress,Planned,Gap"')

    set_column_widths(ws, {"A": 20, "B": 40, "C": 18, "D": 28, "E": 15, "F": 18, "G": 14, "H": 25})
    ws.freeze_panes = "A4"
    logger.info("Created Allocation_Process sheet")


def create_change_management_sheet(wb: Workbook):
    ws = wb.create_sheet("Change_Management")

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Credential Change Management"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["Change Type", "Trigger", "Process Steps", "Verification", "SLA", "Notification", "Status", "Notes"]
    create_header_row(ws, 3, headers)

    changes = [
        ("User-Initiated Change", "User requests password change", "1. Authenticate, 2. Enter new password, 3. Confirm", "Current password verified", "Immediate", "Email confirmation", "", ""),
        ("Scheduled Rotation", "Password expiration reached", "1. Prompt at login, 2. Enter new password, 3. Confirm", "Current password verified", "Before expiry", "7-day advance warning", "", ""),
        ("Administrative Reset", "User requests reset via helpdesk", "1. Identity verification, 2. Generate temp password, 3. Secure delivery", "Security questions/callback", "15 minutes", "Email to user", "", ""),
        ("Security Incident", "Suspected credential compromise", "1. Force immediate change, 2. Terminate sessions, 3. Investigate", "Admin authorization", "Immediate", "User + security team", "", ""),
        ("Role Change", "User changes job role/department", "1. Review access, 2. Revoke old, 3. Grant new", "Manager approval", "1 business day", "User + managers", "", ""),
        ("MFA Device Change", "User changes MFA device", "1. Identity verification, 2. Remove old device, 3. Enroll new", "In-person or secure callback", "Same day", "Email confirmation", "", ""),
        ("Token Refresh", "API key/service account rotation", "1. Generate new key, 2. Update applications, 3. Revoke old", "Change approval", "Scheduled window", "Application owners", "", ""),
        ("Certificate Renewal", "Certificate approaching expiry", "1. Generate CSR, 2. Issue cert, 3. Deploy, 4. Verify", "Auto or manual trigger", "7 days before expiry", "System owners", "", ""),
        ("Emergency Access", "Break-glass credential use", "1. Record justification, 2. Grant access, 3. Audit review", "Dual authorization", "Immediate", "Security team", "", ""),
    ]

    row = 4
    for change in changes:
        for col, value in enumerate(change, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [7, 8]:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"G4:G{row+5}", '"Implemented,In Progress,Planned,Gap"')

    set_column_widths(ws, {"A": 22, "B": 30, "C": 45, "D": 25, "E": 15, "F": 22, "G": 14, "H": 25})
    ws.freeze_panes = "A4"
    logger.info("Created Change_Management sheet")


def create_recovery_process_sheet(wb: Workbook):
    ws = wb.create_sheet("Recovery_Process")

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Credential Recovery Processes"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["Recovery Method", "Use Case", "Process Steps", "Identity Verification", "Security Controls", "SLA", "Status", "Notes"]
    create_header_row(ws, 3, headers)

    methods = [
        ("Self-Service Email", "User forgets password, has email access", "1. Request reset, 2. Click email link, 3. Set new password", "Email ownership", "Time-limited token (15 min), single use", "Immediate", "", ""),
        ("Self-Service SMS", "User forgets password, has phone access", "1. Request reset, 2. Enter SMS code, 3. Set new password", "Phone ownership", "Time-limited code (10 min), rate limited", "Immediate", "", ""),
        ("Security Questions", "Backup recovery method", "1. Answer questions, 2. Verify answers, 3. Reset password", "Knowledge-based", "Limited attempts, lockout after failures", "Immediate", "", ""),
        ("MFA Recovery Codes", "User loses MFA device", "1. Enter recovery code, 2. Access account, 3. Re-enroll MFA", "Recovery code possession", "Single-use codes, secure storage reminder", "Immediate", "", ""),
        ("Helpdesk Assisted", "User cannot self-recover", "1. Call helpdesk, 2. Identity verification, 3. Temp password issued", "Photo ID + security questions", "Callback verification, temp password expires", "15 minutes", "", ""),
        ("Manager Verification", "High-security accounts", "1. Manager confirms identity, 2. IT resets, 3. Secure delivery", "Manager attestation", "Out-of-band confirmation required", "1 hour", "", ""),
        ("In-Person Recovery", "Highest security accounts", "1. Visit IT in person, 2. Photo ID check, 3. Reset performed", "Physical ID verification", "No remote recovery allowed", "Same day", "", ""),
        ("Privileged Account", "Admin/service account recovery", "1. Dual authorization, 2. Break-glass procedure, 3. Full audit", "Dual control required", "Emergency access procedure, immediate review", "15 minutes", "", ""),
    ]

    row = 4
    for method in methods:
        for col, value in enumerate(method, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [7, 8]:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"G4:G{row+5}", '"Implemented,In Progress,Planned,Gap"')

    set_column_widths(ws, {"A": 22, "B": 30, "C": 45, "D": 22, "E": 35, "F": 12, "G": 14, "H": 25})
    ws.freeze_panes = "A4"
    logger.info("Created Recovery_Process sheet")


def create_revocation_process_sheet(wb: Workbook):
    ws = wb.create_sheet("Revocation_Process")

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Credential Revocation Process"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["Revocation Trigger", "SLA Requirement", "Actions Required", "Verification", "Systems Affected", "Responsible", "Status", "Notes"]
    create_header_row(ws, 3, headers)

    triggers = [
        ("Voluntary Resignation", "Last day of employment", "Disable account, revoke all access, terminate sessions", "HR notification received", "AD, Email, VPN, Apps, Cloud", "IT Operations", "", ""),
        ("Involuntary Termination", "Immediate (same hour)", "Immediate disable, revoke all, terminate sessions, collect devices", "HR/Legal notification", "All systems", "IT Operations + Security", "", ""),
        ("Contract End", "Contract end date", "Disable account, revoke access, archive data", "Contract manager notification", "All contractor-accessed systems", "IT Operations", "", ""),
        ("Extended Leave", "Start of leave", "Disable account (not delete), preserve access config", "HR notification", "Interactive login systems", "IT Operations", "", ""),
        ("Role Change", "Effective date of change", "Revoke old access, grant new access per new role", "Manager approval", "Role-specific systems", "IT Operations", "", ""),
        ("Security Incident", "Immediate", "Disable account, terminate sessions, preserve for investigation", "Security team authorization", "All systems", "Security Team", "", ""),
        ("Policy Violation", "Per investigation outcome", "Suspend or revoke based on severity", "HR/Legal/Security decision", "Relevant systems", "Security Team", "", ""),
        ("Account Inactivity", "After defined period (90 days)", "Disable account, notify manager", "Automated detection", "All systems", "Automated + IT", "", ""),
        ("Credential Compromise", "Immediate", "Revoke credential, force reset, audit access", "Security team determination", "Affected credential scope", "Security Team", "", ""),
        ("Device Loss/Theft", "Immediate", "Revoke device certificates, remote wipe, force password change", "User/manager report", "Device-bound credentials", "IT Operations", "", ""),
    ]

    row = 4
    for trigger in triggers:
        for col, value in enumerate(trigger, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [7, 8]:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"G4:G{row+5}", '"Implemented,In Progress,Planned,Gap"')

    set_column_widths(ws, {"A": 22, "B": 22, "C": 45, "D": 25, "E": 30, "F": 20, "G": 14, "H": 25})
    ws.freeze_panes = "A4"
    logger.info("Created Revocation_Process sheet")


def create_audit_log_requirements_sheet(wb: Workbook):
    ws = wb.create_sheet("Audit_Log_Requirements")

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Authentication Audit Log Requirements"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["Event Type", "Details to Log", "Retention Period", "Alerting Required", "Review Frequency", "Status", "Notes"]
    create_header_row(ws, 3, headers)

    events = [
        ("Successful Login", "User, timestamp, source IP, device, MFA method", "1 year minimum", "No (baseline)", "Monthly sampling", "", ""),
        ("Failed Login", "User, timestamp, source IP, failure reason", "1 year minimum", "Yes (threshold)", "Real-time monitoring", "", ""),
        ("Account Lockout", "User, timestamp, source IP, lockout reason", "1 year minimum", "Yes (immediate)", "Real-time monitoring", "", ""),
        ("Password Change", "User, timestamp, change type (user/admin)", "1 year minimum", "Admin changes only", "Weekly review", "", ""),
        ("Password Reset", "User, timestamp, reset method, requester", "1 year minimum", "Yes (admin resets)", "Daily review", "", ""),
        ("MFA Enrollment", "User, timestamp, MFA type, device info", "1 year minimum", "Yes", "Weekly review", "", ""),
        ("MFA Failure", "User, timestamp, failure reason", "1 year minimum", "Yes (threshold)", "Real-time monitoring", "", ""),
        ("Session Start/End", "User, timestamp, duration, source IP", "1 year minimum", "Anomalies only", "Monthly review", "", ""),
        ("Privilege Escalation", "User, timestamp, privilege granted, grantor", "2 years minimum", "Yes (immediate)", "Real-time monitoring", "", ""),
        ("Account Creation", "New user, timestamp, creator, initial access", "2 years minimum", "Yes", "Daily review", "", ""),
        ("Account Deletion", "User, timestamp, deleter, reason", "2 years minimum", "Yes", "Daily review", "", ""),
        ("Access Revocation", "User, timestamp, revoker, access revoked", "2 years minimum", "Yes", "Daily review", "", ""),
        ("Service Account Access", "Account, timestamp, action, target system", "2 years minimum", "Anomalies", "Weekly review", "", ""),
        ("Emergency Access", "User, timestamp, justification, approver", "5 years minimum", "Yes (immediate)", "Same-day review", "", ""),
    ]

    row = 4
    for event in events:
        for col, value in enumerate(event, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [6, 7]:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"F4:F{row+5}", '"Implemented,In Progress,Planned,Gap"')

    set_column_widths(ws, {"A": 22, "B": 40, "C": 18, "D": 20, "E": 20, "F": 14, "G": 25})
    ws.freeze_panes = "A4"
    logger.info("Created Audit_Log_Requirements sheet")


def create_evidence_register_sheet(wb: Workbook):
    ws = wb.create_sheet("Evidence_Register")

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Evidence Register - Credential Lifecycle"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["Evidence ID", "Evidence Type", "Description", "Related Process", "Location/Link", "Date Collected", "Collected By", "Status"]
    create_header_row(ws, 3, headers)

    evidence = [
        ("EV-517-LC-001", "Procedure Document", "User provisioning procedure", "Allocation_Process", "", "", "", ""),
        ("EV-517-LC-002", "System Config", "Password reset workflow configuration", "Change_Management", "", "", "", ""),
        ("EV-517-LC-003", "Procedure Document", "Account recovery procedures", "Recovery_Process", "", "", "", ""),
        ("EV-517-LC-004", "Procedure Document", "Offboarding checklist", "Revocation_Process", "", "", "", ""),
        ("EV-517-LC-005", "Audit Log Sample", "30-day authentication event log", "Audit_Log_Requirements", "", "", "", ""),
    ]

    row = 4
    for item in evidence:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [5, 6, 7, 8]:
                cell.fill = INPUT_FILL
        row += 1

    for i in range(10):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"H4:H{row-1}", '"Pending,Collected,Verified,Expired,N/A"')

    set_column_widths(ws, {"A": 16, "B": 18, "C": 35, "D": 22, "E": 28, "F": 15, "G": 16, "H": 12})
    ws.freeze_panes = "A4"
    logger.info("Created Evidence_Register sheet")


def create_approval_signoff_sheet(wb: Workbook):
    ws = wb.create_sheet("Approval_SignOff")

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Process Approval and Sign-Off"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN

    ws.cell(row=3, column=1, value="Document Information").font = BOLD_FONT

    info = [
        ("Document ID:", DOCUMENT_ID, "Version:", "1.0"),
        ("Document Title:", WORKBOOK_NAME, "Date:", GENERATED_DATE),
        ("Control Reference:", CONTROL_ID, "Classification:", "INTERNAL"),
    ]

    row = 4
    for item in info:
        ws.cell(row=row, column=1, value=item[0]).font = BOLD_FONT
        ws.cell(row=row, column=2, value=item[1]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=item[2]).font = BOLD_FONT
        ws.cell(row=row, column=4, value=item[3]).font = NORMAL_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Approval Signatures").font = BOLD_FONT

    row += 1
    headers = ["Role", "Name", "Signature", "Date", "Status", "Comments"]
    create_header_row(ws, row, headers)

    row += 1
    approvers = [
        ("Process Owner", "", "", "", "", ""),
        ("IT Operations Manager", "", "", "", "", ""),
        ("Information Security Officer", "", "", "", "", ""),
    ]

    for approver in approvers:
        for col, value in enumerate(approver, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col > 1:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"E{row-3}:E{row-1}", '"Pending,Approved,Rejected,Deferred"')

    set_column_widths(ws, {"A": 25, "B": 25, "C": 20, "D": 15, "E": 15, "F": 30})
    ws.freeze_panes = "A4"
    logger.info("Created Approval_SignOff sheet")


def main():
    logger.info("=" * 70)
    logger.info(f"{DOCUMENT_ID} {WORKBOOK_NAME} Generator")
    logger.info("=" * 70)

    wb = Workbook()

    create_instructions_sheet(wb)
    create_allocation_process_sheet(wb)
    create_change_management_sheet(wb)
    create_recovery_process_sheet(wb)
    create_revocation_process_sheet(wb)
    create_audit_log_requirements_sheet(wb)
    create_evidence_register_sheet(wb)
    create_approval_signoff_sheet(wb)

    output_path = Path(__file__).parent / OUTPUT_FILENAME
    wb.save(output_path)

    logger.info("=" * 70)
    logger.info(f"SUCCESS: Workbook saved as {OUTPUT_FILENAME}")
    logger.info("=" * 70)

    return 0


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.17 Authentication Information control
# =============================================================================
