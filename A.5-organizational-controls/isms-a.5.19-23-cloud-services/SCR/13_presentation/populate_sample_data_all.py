#!/usr/bin/env python3
"""
ISMS-IMP-A.5.23 - Sample Data Population Script
==============================================
Populates all 5 cloud services workbooks with realistic sample data.

Usage:
    python3 populate_sample_data_all.py

This script:
1. Loads all regenerated workbooks from 90_workbooks/
2. Fills them with consistent, realistic sample data
3. Saves to 90_workbooks/populated/ folder

Sample data represents a typical Swiss mid-sized company using:
- Microsoft 365, Azure, AWS for core cloud services
- Various SaaS tools (Salesforce, ServiceNow, etc.)
- Compliance focus: GDPR, FADP, DORA, NIS2
"""

import os
import logging
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# =============================================================================
# LOGGING SETUP
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# SAMPLE DATA - CONSISTENT ACROSS ALL WORKBOOKS
# =============================================================================

# Cloud Services Inventory (used across S1, S4, S5)
SAMPLE_SERVICES = [
    {
        'name': 'Microsoft 365',
        'type': 'SaaS',
        'provider': 'Microsoft',
        'criticality': 'Critical',
        'data_class': 'Confidential',
        'region': 'Switzerland (Zurich/Geneva)',
        'contract': 'Active',
        'exit_strategy': 'Cloud-to-Cloud',
        'alternative': 'Google Workspace',
        'export_format': 'PST/CSV',
        'export_tested': 'Yes',
        'lock_in': 'Medium',
        'dora_scope': 'In-Scope',
        'nis2_scope': 'Yes',
        'ai_risk': 'Limited Risk',
        'cloud_act': 'Potential Exposure',
    },
    {
        'name': 'Azure Cloud Platform',
        'type': 'IaaS',
        'provider': 'Microsoft',
        'criticality': 'Critical',
        'data_class': 'Restricted',
        'region': 'Switzerland North',
        'contract': 'Active',
        'exit_strategy': 'Hybrid',
        'alternative': 'AWS/GCP',
        'export_format': 'JSON/ARM Templates',
        'export_tested': 'Yes',
        'lock_in': 'High',
        'dora_scope': 'In-Scope',
        'nis2_scope': 'Yes',
        'ai_risk': 'N/A',
        'cloud_act': 'Potential Exposure',
    },
    {
        'name': 'AWS S3 Storage',
        'type': 'Storage',
        'provider': 'Amazon',
        'criticality': 'High',
        'data_class': 'Confidential',
        'region': 'EU (Frankfurt)',
        'contract': 'Active',
        'exit_strategy': 'Cloud-to-Cloud',
        'alternative': 'Azure Blob',
        'export_format': 'Native/S3 API',
        'export_tested': 'Yes',
        'lock_in': 'Low',
        'dora_scope': 'In-Scope',
        'nis2_scope': 'Yes',
        'ai_risk': 'N/A',
        'cloud_act': 'Potential Exposure',
    },
    {
        'name': 'Salesforce CRM',
        'type': 'SaaS',
        'provider': 'Salesforce',
        'criticality': 'High',
        'data_class': 'Confidential',
        'region': 'EU (Frankfurt)',
        'contract': 'Active',
        'exit_strategy': 'Cloud-to-Cloud',
        'alternative': 'Microsoft Dynamics',
        'export_format': 'CSV/Data Loader',
        'export_tested': 'Yes',
        'lock_in': 'Medium',
        'dora_scope': 'Out-of-Scope',
        'nis2_scope': 'No',
        'ai_risk': 'Limited Risk',
        'cloud_act': 'Potential Exposure',
    },
    {
        'name': 'ServiceNow ITSM',
        'type': 'SaaS',
        'provider': 'ServiceNow',
        'criticality': 'High',
        'data_class': 'Internal',
        'region': 'EU (Amsterdam)',
        'contract': 'Active',
        'exit_strategy': 'Cloud-to-Cloud',
        'alternative': 'Jira Service Mgmt',
        'export_format': 'XML/CSV Export',
        'export_tested': 'Yes',
        'lock_in': 'Medium',
        'dora_scope': 'In-Scope',
        'nis2_scope': 'Yes',
        'ai_risk': 'N/A',
        'cloud_act': 'Low Risk',
    },
    {
        'name': 'SAP S/4HANA Cloud',
        'type': 'SaaS',
        'provider': 'SAP',
        'criticality': 'Critical',
        'data_class': 'Restricted',
        'region': 'Switzerland (Zurich)',
        'contract': 'Active',
        'exit_strategy': 'Hybrid',
        'alternative': 'None viable',
        'export_format': 'RFC/IDoc',
        'export_tested': 'Partial',
        'lock_in': 'Critical',
        'dora_scope': 'In-Scope',
        'nis2_scope': 'Yes',
        'ai_risk': 'N/A',
        'cloud_act': 'No Exposure',
    },
    {
        'name': 'Crowdstrike Falcon',
        'type': 'Security',
        'provider': 'Crowdstrike',
        'criticality': 'Critical',
        'data_class': 'Confidential',
        'region': 'EU (Frankfurt)',
        'contract': 'Active',
        'exit_strategy': 'Cloud-to-Cloud',
        'alternative': 'Microsoft Defender',
        'export_format': 'JSON/SIEM',
        'export_tested': 'Yes',
        'lock_in': 'Low',
        'dora_scope': 'In-Scope',
        'nis2_scope': 'Yes',
        'ai_risk': 'Limited Risk',
        'cloud_act': 'Potential Exposure',
    },
    {
        'name': 'Slack Enterprise',
        'type': 'SaaS',
        'provider': 'Salesforce',
        'criticality': 'Medium',
        'data_class': 'Internal',
        'region': 'EU (Frankfurt)',
        'contract': 'Active',
        'exit_strategy': 'Cloud-to-Cloud',
        'alternative': 'Microsoft Teams',
        'export_format': 'JSON Export',
        'export_tested': 'Yes',
        'lock_in': 'Low',
        'dora_scope': 'Out-of-Scope',
        'nis2_scope': 'No',
        'ai_risk': 'N/A',
        'cloud_act': 'Potential Exposure',
    },
]

# Vendor certifications
VENDOR_CERTS = {
    'Microsoft': ['ISO 27001', 'SOC 2 Type II', 'C5', 'FedRAMP High'],
    'Amazon': ['ISO 27001', 'SOC 2 Type II', 'C5', 'FedRAMP Moderate'],
    'Salesforce': ['ISO 27001', 'SOC 2 Type II', 'C5'],
    'ServiceNow': ['ISO 27001', 'SOC 2 Type II'],
    'SAP': ['ISO 27001', 'SOC 2 Type II', 'C5'],
    'Crowdstrike': ['ISO 27001', 'SOC 2 Type II', 'FedRAMP Moderate'],
}


# =============================================================================
# S1 - INVENTORY WORKBOOK POPULATION
# =============================================================================

def populate_s1_inventory(wb):
    """Populate S1 Cloud Service Inventory workbook."""
    logger.info("Populating S1 - Cloud Service Inventory...")

    # Sheet 2: SaaS Services
    ws = wb['2. SaaS Services']
    saas_services = [s for s in SAMPLE_SERVICES if s['type'] == 'SaaS']
    row = 8  # Data starts at row 8
    for svc in saas_services:
        ws[f'A{row}'] = svc['name']
        ws[f'B{row}'] = svc['type']
        ws[f'C{row}'] = svc['provider']
        ws[f'D{row}'] = svc['criticality']
        ws[f'E{row}'] = svc['data_class']
        ws[f'F{row}'] = svc['region']
        ws[f'G{row}'] = svc['contract']
        ws[f'H{row}'] = '✓ Compliant'
        row += 1
    logger.info(f"   [OK] SaaS Services: {len(saas_services)} entries")

    # Sheet 3: IaaS/PaaS Services
    ws = wb['3. IaaS PaaS Services']
    iaas_services = [s for s in SAMPLE_SERVICES if s['type'] in ['IaaS', 'PaaS']]
    row = 8
    for svc in iaas_services:
        ws[f'A{row}'] = svc['name']
        ws[f'B{row}'] = svc['type']
        ws[f'C{row}'] = svc['provider']
        ws[f'D{row}'] = svc['criticality']
        ws[f'E{row}'] = svc['data_class']
        ws[f'F{row}'] = svc['region']
        ws[f'G{row}'] = svc['contract']
        ws[f'H{row}'] = '✓ Compliant'
        row += 1
    logger.info(f"   [OK] IaaS/PaaS Services: {len(iaas_services)} entries")

    # Sheet 4: Exit Strategy
    ws = wb['4. Cloud Security Services']
    row = 4
    for svc in SAMPLE_SERVICES:
        ws[f'A{row}'] = svc['name']
        ws[f'B{row}'] = svc['provider']
        ws[f'C{row}'] = svc['type']
        ws[f'D{row}'] = 'Production'
        ws[f'E{row}'] = '2024-01-15'
        ws[f'F{row}'] = '2027-01-14'
        ws[f'G{row}'] = svc['data_class']
        ws[f'H{row}'] = svc['criticality']
        ws[f'R{row}'] = svc['exit_strategy']
        ws[f'S{row}'] = svc['alternative']
        ws[f'T{row}'] = svc['export_format']
        ws[f'U{row}'] = svc['export_tested']
        ws[f'X{row}'] = svc['lock_in']
        row += 1
    logger.info(f"   [OK] Exit Strategy: {len(SAMPLE_SERVICES)} entries")

    # Sheet 5: Cloud Storage
    ws = wb['5. Cloud Storage Services']
    storage_services = [s for s in SAMPLE_SERVICES if s['type'] == 'Storage']
    row = 8
    for svc in storage_services:
        ws[f'A{row}'] = svc['name']
        ws[f'B{row}'] = svc['type']
        ws[f'C{row}'] = svc['provider']
        ws[f'D{row}'] = svc['criticality']
        ws[f'E{row}'] = svc['data_class']
        ws[f'F{row}'] = svc['region']
        ws[f'G{row}'] = svc['contract']
        ws[f'H{row}'] = '✓ Compliant'
        row += 1
    logger.info(f"   [OK] Cloud Storage: {len(storage_services)} entries")

    # Sheet 6: Data Classification
    ws = wb['6. Data Classification']
    data_categories = [
        ('Customer PII', 'Personal data of customers', 'Restricted', 'Microsoft 365', 'SharePoint', 'Legal', '7 years', 'Yes - Personal', 'EEA Only'),
        ('Employee HR Data', 'Employee records and payroll', 'Restricted', 'SAP S/4HANA', 'SAP Cloud', 'HR Director', '10 years', 'Yes - Special Category', 'None'),
        ('Financial Reports', 'Quarterly and annual financials', 'Confidential', 'SAP S/4HANA', 'SAP Cloud', 'CFO', '10 years', 'No', 'None'),
        ('Sales Data', 'CRM and pipeline data', 'Confidential', 'Salesforce CRM', 'Salesforce', 'Sales Director', '5 years', 'Yes - Personal', 'Adequacy Decision'),
        ('IT Logs', 'System and security logs', 'Internal', 'Azure Cloud', 'Azure Log Analytics', 'CISO', '1 year', 'No', 'EEA Only'),
    ]
    row = 4
    for cat in data_categories:
        ws[f'A{row}'] = f'DC-{row-3:03d}'
        ws[f'B{row}'] = cat[0]
        ws[f'C{row}'] = cat[1]
        ws[f'D{row}'] = cat[2]
        ws[f'E{row}'] = cat[3]
        ws[f'F{row}'] = cat[4]
        ws[f'G{row}'] = cat[5]
        ws[f'H{row}'] = cat[6]
        ws[f'I{row}'] = cat[7]
        ws[f'J{row}'] = cat[8]
        row += 1
    logger.info(f"   [OK] Data Classification: {len(data_categories)} entries")

    # Sheet 7: Service Criticality
    ws = wb['7. Service Criticality']
    row = 4
    for svc in SAMPLE_SERVICES:
        rto = 4 if svc['criticality'] == 'Critical' else (8 if svc['criticality'] == 'High' else 24)
        rpo = 1 if svc['criticality'] == 'Critical' else (4 if svc['criticality'] == 'High' else 24)
        ws[f'A{row}'] = f'SC-{row-3:03d}'
        ws[f'B{row}'] = svc['name']
        ws[f'C{row}'] = svc['type']
        ws[f'D{row}'] = 'Core Business Operations'
        ws[f'E{row}'] = svc['criticality']
        ws[f'F{row}'] = rto
        ws[f'G{row}'] = rpo
        ws[f'H{row}'] = rto * 3
        ws[f'I{row}'] = 'No' if svc['alternative'] != 'None viable' else 'Yes'
        ws[f'J{row}'] = 'Yes' if svc['alternative'] != 'None viable' else 'No'
        ws[f'K{row}'] = svc['dora_scope']
        ws[f'L{row}'] = f"P1 - Critical" if svc['criticality'] == 'Critical' else f"P2 - High"
        row += 1
    logger.info(f"   [OK] Service Criticality: {len(SAMPLE_SERVICES)} entries")

    return wb


# =============================================================================
# S2 - VENDOR DUE DILIGENCE POPULATION
# =============================================================================

def populate_s2_vendor_dd(wb):
    """Populate S2 Vendor Due Diligence workbook."""
    logger.info("Populating S2 - Vendor Due Diligence...")

    # Sheet 2: Security Certifications
    ws = wb['2. Security Certifications']
    row = 4
    for vendor, certs in VENDOR_CERTS.items():
        ws[f'A{row}'] = vendor
        ws[f'B{row}'] = ', '.join(certs)
        ws[f'C{row}'] = 'Yes' if 'ISO 27001' in certs else 'No'
        ws[f'D{row}'] = 'Yes' if 'SOC 2' in certs else 'No'
        ws[f'E{row}'] = 'Yes' if 'C5' in certs else 'No'
        ws[f'F{row}'] = '2024-06-15'
        ws[f'G{row}'] = '2025-06-14'
        ws[f'H{row}'] = '✓ Compliant'
        row += 1
    logger.info(f"   [OK] Security Certifications: {len(VENDOR_CERTS)} vendors")

    # Sheet 3: Contract Terms
    ws = wb['3. Contract Terms']
    row = 4
    for svc in SAMPLE_SERVICES:
        ws[f'A{row}'] = svc['provider']
        ws[f'B{row}'] = svc['name']
        ws[f'C{row}'] = '2024-01-15'
        ws[f'D{row}'] = '2027-01-14'
        ws[f'E{row}'] = 'Yes'  # DPA signed
        ws[f'F{row}'] = 'Yes'  # SCCs
        ws[f'G{row}'] = 'Yes'  # DORA Art 30
        ws[f'H{row}'] = '30 days'
        ws[f'I{row}'] = 'CHF 5M'
        ws[f'J{row}'] = '✓ Compliant'
        row += 1
    logger.info(f"   [OK] Contract Terms: {len(SAMPLE_SERVICES)} contracts")

    # Sheet 5: Data Sovereignty
    ws = wb['5. Data Sovereignty']
    row = 4
    for svc in SAMPLE_SERVICES:
        ws[f'A{row}'] = svc['provider']
        ws[f'B{row}'] = svc['name']
        ws[f'C{row}'] = svc['region']
        ws[f'D{row}'] = 'Yes' if 'Switzerland' in svc['region'] else 'No'
        ws[f'E{row}'] = 'Yes'  # GDPR compliant
        ws[f'F{row}'] = 'Yes'  # nFADP compliant
        ws[f'G{row}'] = 'SCCs' if 'EU' in svc['region'] else 'N/A'
        ws[f'H{row}'] = '✓ Compliant'
        row += 1
    logger.info(f"   [OK] Data Sovereignty: {len(SAMPLE_SERVICES)} entries")

    # Sheet 7: Jurisdictional Risk (CLOUD Act)
    ws = wb['7. Jurisdictional Risk']
    row = 4
    for svc in SAMPLE_SERVICES:
        us_nexus = svc['provider'] in ['Microsoft', 'Amazon', 'Salesforce', 'Crowdstrike', 'ServiceNow']
        ws[f'A{row}'] = svc['provider']
        ws[f'B{row}'] = svc['name']
        ws[f'C{row}'] = 'Yes' if us_nexus else 'No'
        ws[f'D{row}'] = 'US' if us_nexus else 'Germany'
        ws[f'E{row}'] = svc['cloud_act']
        ws[f'F{row}'] = 'Technical measures' if svc['cloud_act'] == 'Potential Exposure' else 'N/A'
        ws[f'G{row}'] = '⚠️ Review' if svc['cloud_act'] == 'Potential Exposure' else '✓ OK'
        row += 1
    logger.info(f"   [OK] Jurisdictional Risk: {len(SAMPLE_SERVICES)} entries")

    return wb


# =============================================================================
# S3 - SECURE CONFIGURATION POPULATION
# =============================================================================

def populate_s3_secure_config(wb):
    """Populate S3 Secure Configuration workbook."""
    logger.info("Populating S3 - Secure Configuration...")

    # Sheet 2: Configuration Baseline
    ws = wb['2. Configuration Baseline']
    configs = [
        ('Azure MFA', 'Authentication', 'Enabled', 'Yes', 'Yes', 'DORA Compliant'),
        ('AWS IAM Policies', 'Access Control', 'Least Privilege', 'Yes', 'Yes', 'DORA Compliant'),
        ('M365 Conditional Access', 'Authentication', 'Risk-based', 'Yes', 'Yes', 'DORA Compliant'),
        ('Salesforce SSO', 'Authentication', 'SAML/OIDC', 'Yes', 'Yes', 'NIS2 Compliant'),
        ('S3 Bucket Encryption', 'Encryption', 'AES-256', 'Yes', 'Yes', 'DORA Compliant'),
    ]
    row = 4
    for cfg in configs:
        ws[f'A{row}'] = cfg[0]
        ws[f'B{row}'] = cfg[1]
        ws[f'C{row}'] = cfg[2]
        ws[f'D{row}'] = cfg[3]
        ws[f'E{row}'] = cfg[4]
        ws[f'F{row}'] = cfg[5]
        ws[f'G{row}'] = '✓ Compliant'
        row += 1
    logger.info(f"   [OK] Configuration Baseline: {len(configs)} configs")

    # Sheet 3: Access Control Setup
    ws = wb['3. Access Control Setup']
    row = 4
    for svc in SAMPLE_SERVICES[:5]:
        ws[f'A{row}'] = svc['name']
        ws[f'B{row}'] = 'Yes'  # SSO Enabled
        ws[f'C{row}'] = 'Yes'  # MFA Enabled
        ws[f'D{row}'] = 'Yes'  # RBAC
        ws[f'E{row}'] = 'Yes'  # JIT Access
        ws[f'F{row}'] = '90 days'  # Access review
        ws[f'G{row}'] = '✓ Compliant'
        row += 1
    logger.info(f"   [OK] Access Control Setup: 5 services")

    # Sheet 5: Encryption Configuration
    ws = wb['5. Encryption Configuration']
    row = 4
    for svc in SAMPLE_SERVICES[:5]:
        ws[f'A{row}'] = svc['name']
        ws[f'B{row}'] = 'AES-256'  # At rest
        ws[f'C{row}'] = 'TLS 1.3'  # In transit
        ws[f'D{row}'] = 'Customer-managed' if svc['criticality'] == 'Critical' else 'Provider-managed'
        ws[f'E{row}'] = 'Azure Key Vault' if svc['provider'] == 'Microsoft' else 'AWS KMS'
        ws[f'F{row}'] = '✓ Compliant'
        row += 1
    logger.info(f"   [OK] Encryption Configuration: 5 services")

    return wb


# =============================================================================
# S4 - GOVERNANCE POPULATION
# =============================================================================

def populate_s4_governance(wb):
    """Populate S4 Governance workbook."""
    logger.info("Populating S4 - Governance...")

    # Sheet 2: Access Review
    ws = wb['2. Access Review']
    row = 4
    for svc in SAMPLE_SERVICES[:5]:
        ws[f'A{row}'] = svc['name']
        ws[f'B{row}'] = '2024-12-01'
        ws[f'C{row}'] = '2025-03-01'
        ws[f'D{row}'] = 'IT Security'
        ws[f'E{row}'] = 'Completed'
        ws[f'F{row}'] = '2 users removed'
        ws[f'G{row}'] = '✓ Compliant'
        row += 1
    logger.info(f"   [OK] Access Review: 5 services")

    # Sheet 4: Incident Management
    ws = wb['4. Incident Management']
    incidents = [
        ('INC-001', 'M365 Phishing Attempt', 'Medium', 'Detected', 'Security Team', 'DORA Reported'),
        ('INC-002', 'AWS S3 Misconfiguration', 'High', 'Resolved', 'Cloud Team', 'DORA Reported'),
        ('INC-003', 'Salesforce Login Anomaly', 'Low', 'Closed', 'SOC', 'Not Applicable'),
    ]
    row = 4
    for inc in incidents:
        ws[f'A{row}'] = inc[0]
        ws[f'B{row}'] = inc[1]
        ws[f'C{row}'] = inc[2]
        ws[f'D{row}'] = inc[3]
        ws[f'E{row}'] = inc[4]
        ws[f'F{row}'] = inc[5]
        row += 1
    logger.info(f"   [OK] Incident Management: {len(incidents)} incidents")

    # Sheet 7: Exit Strategy Review
    ws = wb['7. Exit Strategy Review']
    today = datetime.now()
    row = 4
    for svc in SAMPLE_SERVICES:
        last_review = today - timedelta(days=180)
        next_review = today + timedelta(days=185)
        poc_required = 'Yes (Critical)' if svc['criticality'] == 'Critical' else 'No (Not Critical)'
        ws[f'A{row}'] = svc['name']
        ws[f'B{row}'] = svc['provider']
        ws[f'C{row}'] = svc['criticality']
        ws[f'D{row}'] = svc['exit_strategy']
        ws[f'E{row}'] = last_review.strftime('%Y-%m-%d')
        ws[f'F{row}'] = next_review.strftime('%Y-%m-%d')
        ws[f'G{row}'] = 'Current'
        ws[f'H{row}'] = poc_required
        if svc['criticality'] == 'Critical':
            ws[f'I{row}'] = (today - timedelta(days=60)).strftime('%Y-%m-%d')
            ws[f'J{row}'] = 'Pass'
            ws[f'K{row}'] = (today + timedelta(days=305)).strftime('%Y-%m-%d')
        else:
            ws[f'I{row}'] = ''
            ws[f'J{row}'] = 'Not Applicable'
            ws[f'K{row}'] = ''
        ws[f'L{row}'] = 'No'
        ws[f'M{row}'] = 'J. Smith'
        ws[f'N{row}'] = 'Annual review completed'
        row += 1
    logger.info(f"   [OK] Exit Strategy Review: {len(SAMPLE_SERVICES)} services")

    return wb


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main execution - populate all workbooks with sample data."""
    logger.info("=" * 70)
    logger.info("ISMS-IMP-A.5.23 - Sample Data Population Script")
    logger.info("=" * 70)

    base_path = '/Users/greg/Files/Factory/factory_claude_ai/10-isms-scr-base/isms-a.5.19-23-cloud-services/90_workbooks'
    output_path = f'{base_path}/populated'

    # Create output directory
    os.makedirs(output_path, exist_ok=True)

    # Workbook mapping
    workbooks = {
        'S1': ('ISMS-IMP-A.5.23.S1_Inventory_20260202.xlsx', populate_s1_inventory),
        'S2': ('ISMS-IMP-A.5.23.S2_VendorDD_20260202.xlsx', populate_s2_vendor_dd),
        'S3': ('ISMS-IMP-A.5.23.S3_SecureConfig_20260202.xlsx', populate_s3_secure_config),
        'S4': ('ISMS-IMP-A.5.23.S4_Governance_20260202.xlsx', populate_s4_governance),
    }

    for wb_id, (filename, populate_func) in workbooks.items():
        filepath = f'{base_path}/{filename}'

        if not os.path.exists(filepath):
            logger.warning(f"[SKIP] {filename} not found")
            continue

        logger.info(f"\n[{wb_id}] Loading {filename}...")
        wb = load_workbook(filepath)

        # Populate with sample data
        wb = populate_func(wb)

        # Save to output folder
        output_file = f'{output_path}/{filename.replace(".xlsx", "_SAMPLE.xlsx")}'
        wb.save(output_file)
        logger.info(f"[{wb_id}] Saved: {output_file.split('/')[-1]}")

    # Note about S5 - it reads from other workbooks
    logger.info("\n" + "=" * 70)
    logger.info("✓ SAMPLE DATA POPULATION COMPLETE")
    logger.info("=" * 70)
    logger.info(f"\nOutput folder: {output_path}")
    logger.info("\nNote: S5 Dashboard regenerates from S1-S4 data.")
    logger.info("      Run generate_reg_a523_5_compliance_dashboard.py after populating S1-S4.")


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - UTILITY SCRIPT
# =============================================================================

