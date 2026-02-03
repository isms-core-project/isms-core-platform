#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
Excel Workbook Sanity Checker - ISMS A.5.24-28 Incident Management Lifecycle
================================================================================

Domain-specific diagnostic utility for A.5.24-28 Incident Management Lifecycle
assessment workbooks covering all 6 assessment domains (S1-S6).

**Purpose:**
Validates workbook structure, formulas, and data validations specific to
incident management assessment requirements across all five assessment domains.

**Usage:**
    python3 sanity_check_a524_28.py ISMS-IMP-A.5.24-28.S1_Framework_Assessment_YYYYMMDD.xlsx
    python3 sanity_check_a524_28.py ISMS-IMP-A.5.24-28.S2_Detection_Classification_YYYYMMDD.xlsx
    python3 sanity_check_a524_28.py ISMS-IMP-A.5.24-28.S3_Response_Capabilities_YYYYMMDD.xlsx
    python3 sanity_check_a524_28.py ISMS-IMP-A.5.24-28.S4_Forensic_Evidence_YYYYMMDD.xlsx
    python3 sanity_check_a524_28.py ISMS-IMP-A.5.24-28.S5_Learning_Improvement_YYYYMMDD.xlsx
    python3 sanity_check_a524_28.py ISMS-IMP-A.5.24-28.S6_Compliance_Dashboard_YYYYMMDD.xlsx
    python3 sanity_check_a524_28.py --all ../90_workbooks/

**Workbook Structures:**
- S1 Framework Assessment: 10 sheets
- S2 Detection & Classification: 9 sheets
- S3 Response Capabilities: 11 sheets
- S4 Forensic Evidence: 10 sheets
- S5 Learning & Improvement: 10 sheets
- S6 Compliance Dashboard: 7 sheets (consolidates S1-S5 metrics)

**Checks Performed:**
- Sheet structure validation
- Data validation dropdowns
- Formula integrity (balanced parentheses, quotes)
- Conditional formatting presence
- Evidence register linkage
- Gap analysis structure

**Exit Codes:**
    0 = All checks passed
    1 = Warnings detected (workbook usable)
    2 = Critical errors detected (regenerate recommended)

Control Reference: ISO/IEC 27001:2022 Annex A Controls A.5.24-28
Version: 1.0
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import os
import sys
from glob import glob
from openpyxl import load_workbook

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# WORKBOOK STRUCTURE DEFINITIONS
# =============================================================================

WORKBOOK_DEFINITIONS = {
    "S1": {
        "name": "Framework Assessment",
        "pattern": "ISMS-IMP-A.5.24-28.S1",
        "expected_sheets": [
            "Instructions & Legend",
            "Governance Assessment",
            "Organizational Structure",
            "Training & Competency",
            "Tools & Technology",
            "Integration Assessment",
            "Gap Analysis",
            "Evidence Register",
            "Dashboard",
            "Approval Sign-Off"
        ],
        "validations": {
            "Governance Assessment": {
                "expected_lists": ["Yes", "No", "Partial", "N/A"],
                "maturity_levels": ["Level 1", "Level 2", "Level 3", "Level 4", "Level 5"]
            },
            "Organizational Structure": {
                "expected_lists": ["Dedicated", "Hybrid", "Virtual", "Outsourced"]
            },
            "Gap Analysis": {
                "severities": ["Critical", "High", "Medium", "Low"],
                "statuses": ["Open", "In Progress", "Planned", "Resolved"]
            }
        }
    },
    "S2": {
        "name": "Detection & Classification",
        "pattern": "ISMS-IMP-A.5.24-28.S2",
        "expected_sheets": [
            "Instructions & Legend",
            "Detection Mechanisms",
            "Alert Handling",
            "Classification & Severity",
            "Detection Effectiveness",
            "Gap Analysis",
            "Evidence Register",
            "Dashboard",
            "Approval Sign-Off"
        ],
        "validations": {
            "Detection Mechanisms": {
                "detection_types": ["SIEM", "EDR", "IDS/IPS", "NDR", "User Reporting", "UEBA"],
                "coverage_levels": ["Full", "Partial", "None", "Unknown"]
            },
            "Classification & Severity": {
                "severities": ["Critical", "High", "Medium", "Low", "Informational"],
                "categories": ["Malware", "Phishing", "Data Breach", "DoS", "Insider Threat"]
            },
            "Gap Analysis": {
                "severities": ["Critical", "High", "Medium", "Low"],
                "statuses": ["Open", "In Progress", "Planned", "Resolved"]
            }
        }
    },
    "S3": {
        "name": "Response Capabilities",
        "pattern": "ISMS-IMP-A.5.24-28.S3",
        "expected_sheets": [
            "Instructions & Legend",
            "Containment Capabilities",
            "Eradication & Remediation",
            "Recovery & Restoration",
            "Communication",
            "Resources & Authority",
            "Playbook Effectiveness",
            "Gap Analysis",
            "Evidence Register",
            "Dashboard",
            "Approval Sign-Off"
        ],
        "validations": {
            "Containment Capabilities": {
                "containment_types": ["Network Isolation", "Endpoint Quarantine", "Account Disable", "Service Shutdown"],
                "effectiveness": ["Highly Effective", "Effective", "Partially Effective", "Ineffective"]
            },
            "Resources & Authority": {
                "authority_levels": ["Immediate", "Manager Approval", "Executive Approval", "Legal Approval"]
            },
            "Gap Analysis": {
                "severities": ["Critical", "High", "Medium", "Low"],
                "statuses": ["Open", "In Progress", "Planned", "Resolved"]
            }
        }
    },
    "S4": {
        "name": "Forensic Evidence",
        "pattern": "ISMS-IMP-A.5.24-28.S4",
        "expected_sheets": [
            "Instructions & Legend",
            "Evidence Collection",
            "Chain of Custody",
            "Forensic Analysis",
            "Storage & Retention",
            "Legal & Regulatory",
            "Gap Analysis",
            "Evidence Register",
            "Dashboard",
            "Approval Sign-Off"
        ],
        "validations": {
            "Evidence Collection": {
                "evidence_types": ["Disk Image", "Memory Dump", "Network Capture", "Log Files", "Volatile Data"],
                "collection_status": ["Complete", "Partial", "Not Collected", "N/A"]
            },
            "Chain of Custody": {
                "integrity_status": ["Verified", "Unverified", "Compromised", "Unknown"]
            },
            "Legal & Regulatory": {
                "admissibility": ["Admissible", "Likely Admissible", "Questionable", "Inadmissible"]
            },
            "Gap Analysis": {
                "severities": ["Critical", "High", "Medium", "Low"],
                "statuses": ["Open", "In Progress", "Planned", "Resolved"]
            }
        }
    },
    "S5": {
        "name": "Learning & Improvement",
        "pattern": "ISMS-IMP-A.5.24-28.S5",
        "expected_sheets": [
            "Instructions & Legend",
            "PIR Process",
            "Root Cause Analysis",
            "Lessons Learned",
            "Control Improvements",
            "Trend Analysis",
            "Gap Analysis",
            "Evidence Register",
            "Summary Dashboard",
            "Approval Sign-Off"
        ],
        "validations": {
            "PIR Process": {
                "pir_quality": ["Level 1", "Level 2", "Level 3", "Level 4", "Level 5"],
                "completion_status": ["Complete", "In Progress", "Overdue", "Not Started"]
            },
            "Root Cause Analysis": {
                "rca_depth": ["Technical", "Procedural", "Systemic", "Combined"],
                "verification": ["Verified", "Partially Verified", "Unverified"]
            },
            "Control Improvements": {
                "action_status": ["Open", "In Progress", "Completed", "Deferred", "Cancelled"]
            },
            "Gap Analysis": {
                "severities": ["Critical", "High", "Medium", "Low"],
                "statuses": ["Open", "In Progress", "Planned", "Resolved"]
            }
        }
    },
    "S6": {
        "name": "Compliance Dashboard",
        "pattern": "ISMS-IMP-A.5.24-28.S6",
        "expected_sheets": [
            "Instructions",
            "Executive_Summary",
            "Compliance_Scores",
            "Gap_Analysis",
            "Incident_KPIs",
            "Evidence_Register",
            "Approval_Sign_Off"
        ],
        "validations": {
            "Compliance_Scores": {
                "compliance_status": ["Compliant", "Partial", "Non-Compliant", "Not Assessed"],
                "maturity_level": ["Level 1", "Level 2", "Level 3", "Level 4", "Level 5"]
            },
            "Gap_Analysis": {
                "severities": ["Critical", "High", "Medium", "Low"],
                "statuses": ["Open", "In Progress", "Planned", "Resolved", "Closed"]
            },
            "Incident_KPIs": {
                "kpi_status": ["On Target", "At Risk", "Off Target", "Not Measured"]
            },
            "Evidence_Register": {
                "verification_status": ["Verified", "Pending", "Not Available", "Expired"]
            }
        }
    }
}


def detect_workbook_type(filename):
    """Detect which workbook type based on filename."""
    for wb_type, definition in WORKBOOK_DEFINITIONS.items():
        if definition["pattern"] in filename:
            return wb_type, definition
    return None, None


def check_sheet_structure(wb, expected_sheets):
    """Check if all expected sheets are present."""
    issues = []
    warnings = []

    found_sheets = wb.sheetnames
    missing_sheets = [s for s in expected_sheets if s not in found_sheets]
    extra_sheets = [s for s in found_sheets if s not in expected_sheets]

    if missing_sheets:
        for sheet in missing_sheets:
            issues.append(f"Missing required sheet: '{sheet}'")

    if extra_sheets:
        for sheet in extra_sheets:
            warnings.append(f"Unexpected sheet found: '{sheet}'")

    return issues, warnings, missing_sheets, extra_sheets


def check_data_validations(ws, validation_config):
    """Check for expected data validations in a worksheet."""
    warnings = []
    found_validations = {}

    if not hasattr(ws, 'data_validations') or not ws.data_validations:
        return warnings, found_validations

    for dv in ws.data_validations.dataValidation:
        if dv.formula1:
            formula = str(dv.formula1)
            for key, values in validation_config.items():
                if isinstance(values, list):
                    if any(v in formula for v in values):
                        found_validations[key] = True

    # Check for missing validations
    for key in validation_config.keys():
        if key not in found_validations:
            warnings.append(f"Validation not detected: {key}")

    return warnings, found_validations


def check_formula_integrity(wb):
    """Check formulas for syntax errors across all sheets."""
    issues = []
    formula_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
                    formula = cell.value

                    # Check for unbalanced parentheses
                    if formula.count('(') != formula.count(')'):
                        issues.append(f"{sheet_name}!{cell.coordinate}: Unbalanced parentheses")

                    # Check for unbalanced quotes
                    if formula.count('"') % 2 != 0:
                        issues.append(f"{sheet_name}!{cell.coordinate}: Unbalanced quotes")

    return issues, formula_count


def check_conditional_formatting(wb):
    """Check for conditional formatting presence."""
    total_cf_rules = 0
    sheets_with_cf = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
            cf_count = len(ws.conditional_formatting._cf_rules)
            if cf_count > 0:
                sheets_with_cf += 1
                total_cf_rules += cf_count

    return total_cf_rules, sheets_with_cf


def check_evidence_register(wb):
    """Check Evidence Register structure and content."""
    warnings = []

    evidence_sheet_names = ["Evidence Register", "Evidence_Register"]
    ws = None

    for name in evidence_sheet_names:
        if name in wb.sheetnames:
            ws = wb[name]
            break

    if ws is None:
        return warnings, False

    has_content = False
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=10):
        for cell in row:
            if cell.value:
                has_content = True
                break
        if has_content:
            break

    if not has_content:
        warnings.append("Evidence Register appears empty")

    # Check for data validations
    if not hasattr(ws, 'data_validations') or not ws.data_validations:
        warnings.append("No data validations in Evidence Register")

    return warnings, has_content


def check_gap_analysis(wb):
    """Check Gap Analysis structure and validations."""
    warnings = []

    gap_sheet_names = ["Gap Analysis", "Gap_Analysis"]
    ws = None

    for name in gap_sheet_names:
        if name in wb.sheetnames:
            ws = wb[name]
            break

    if ws is None:
        return warnings

    expected_severities = ['Critical', 'High', 'Medium', 'Low']
    expected_statuses = ['Open', 'In Progress', 'Planned', 'Resolved']

    if hasattr(ws, 'data_validations') and ws.data_validations:
        severity_found = False
        status_found = False

        for dv in ws.data_validations.dataValidation:
            if dv.formula1:
                formula = str(dv.formula1)
                if any(s in formula for s in expected_severities):
                    severity_found = True
                if any(s in formula for s in expected_statuses):
                    status_found = True

        if not severity_found:
            warnings.append("Gap Analysis: Severity validation not detected")
        if not status_found:
            warnings.append("Gap Analysis: Status validation not detected")
    else:
        warnings.append("Gap Analysis: No data validations found")

    return warnings


def check_workbook(filename, wb_type=None, definition=None):
    """Main validation function for a workbook."""

    # Auto-detect workbook type if not provided
    if wb_type is None:
        wb_type, definition = detect_workbook_type(filename)

    if definition is None:
        logger.error(f"Cannot determine workbook type for: {filename}")
        logger.error("Expected pattern: ISMS-IMP-A.5.24-28.S1 through S5")
        return 2

    print("=" * 80)
    print(f"A.5.24-28 INCIDENT MANAGEMENT - {definition['name'].upper()} - SANITY CHECK")
    print("=" * 80)
    print(f"File: {filename}")
    print(f"Workbook Type: {wb_type}\n")

    issues = []
    warnings = []

    # Load workbook
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"[PASS] Workbook loaded: {len(wb.sheetnames)} sheets\n")
    except Exception as e:
        print(f"[CRITICAL] Cannot load workbook: {e}")
        return 2

    # ========================================================================
    # CHECK 1: SHEET STRUCTURE
    # ========================================================================
    print("=" * 80)
    print("CHECK 1: SHEET STRUCTURE")
    print("=" * 80)

    sheet_issues, sheet_warnings, missing, extra = check_sheet_structure(
        wb, definition["expected_sheets"]
    )
    issues.extend(sheet_issues)
    warnings.extend(sheet_warnings)

    print(f"  Expected: {len(definition['expected_sheets'])} sheets")
    print(f"  Found: {len(wb.sheetnames)} sheets")

    if not missing and not extra:
        print("  [PASS] All expected sheets present, no extras")
    else:
        if missing:
            for sheet in missing:
                print(f"  [CRITICAL] Missing: '{sheet}'")
        if extra:
            for sheet in extra:
                print(f"  [WARNING] Extra: '{sheet}'")

    # ========================================================================
    # CHECK 2: DATA VALIDATIONS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 2: DATA VALIDATIONS")
    print("=" * 80)

    total_validations = 0
    for sheet_name, validation_config in definition.get("validations", {}).items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            val_warnings, found = check_data_validations(ws, validation_config)

            if hasattr(ws, 'data_validations') and ws.data_validations:
                dv_count = len(ws.data_validations.dataValidation)
                total_validations += dv_count
                print(f"  {sheet_name}: {dv_count} validations")

                for key in found:
                    print(f"    [PASS] {key} validation detected")
            else:
                warnings.append(f"{sheet_name}: No data validations found")
                print(f"  {sheet_name}: [WARNING] No validations")

            warnings.extend([f"{sheet_name}: {w}" for w in val_warnings])

    if total_validations > 0:
        print(f"\n  Total data validations: {total_validations}")

    # ========================================================================
    # CHECK 3: FORMULA INTEGRITY
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 3: FORMULA INTEGRITY")
    print("=" * 80)

    formula_issues, formula_count = check_formula_integrity(wb)

    print(f"  Total formulas found: {formula_count}")

    if formula_issues:
        for issue in formula_issues:
            issues.append(f"Formula error: {issue}")
            print(f"  [CRITICAL] {issue}")
    else:
        print("  [PASS] No formula syntax errors detected")

    # ========================================================================
    # CHECK 4: CONDITIONAL FORMATTING
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 4: CONDITIONAL FORMATTING")
    print("=" * 80)

    cf_rules, cf_sheets = check_conditional_formatting(wb)

    print(f"  Total conditional formatting rules: {cf_rules}")
    print(f"  Sheets with conditional formatting: {cf_sheets}")

    if cf_rules > 0:
        print("  [PASS] Conditional formatting present")
    else:
        warnings.append("No conditional formatting detected")
        print("  [WARNING] No conditional formatting detected")

    # ========================================================================
    # CHECK 5: EVIDENCE REGISTER
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 5: EVIDENCE REGISTER")
    print("=" * 80)

    ev_warnings, has_content = check_evidence_register(wb)

    if has_content:
        print("  [PASS] Evidence Register has content/structure")
    else:
        print("  [WARNING] Evidence Register may be empty")

    warnings.extend([f"Evidence Register: {w}" for w in ev_warnings])

    # ========================================================================
    # CHECK 6: GAP ANALYSIS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 6: GAP ANALYSIS")
    print("=" * 80)

    gap_warnings = check_gap_analysis(wb)

    if not gap_warnings:
        print("  [PASS] Gap Analysis structure validated")
    else:
        for w in gap_warnings:
            print(f"  [WARNING] {w}")

    warnings.extend(gap_warnings)

    # ========================================================================
    # SUMMARY REPORT
    # ========================================================================
    print("\n" + "=" * 80)
    print("VALIDATION SUMMARY")
    print("=" * 80)

    if issues:
        print(f"\n[CRITICAL] ISSUES: {len(issues)}")
        for issue in issues:
            print(f"  - {issue}")

    if warnings:
        print(f"\n[WARNING] WARNINGS: {len(warnings)}")
        for warning in warnings:
            print(f"  - {warning}")

    if not issues and not warnings:
        print(f"\n[PASS] ALL CHECKS PASSED")
        print(f"\nThe A.5.24-28 {definition['name']} workbook appears healthy.")
        print("Ready for incident management assessment activities.")
    else:
        print("\n" + "=" * 80)
        print("RECOMMENDATIONS:")
        print("=" * 80)

        if missing:
            print(f"\n- Regenerate workbook with generate_a524_28_{wb_type.lower()}*.py")
            print("  to ensure all required sheets are present")

        if warnings:
            print("\n- Review warnings - most are informational")
            print("- Data validations improve data quality but are optional")
            print("- Ensure assessment data is current and accurate")

    print("\n" + "=" * 80)

    # Return exit code
    if issues:
        return 2
    elif warnings:
        return 1
    else:
        return 0


def check_all_workbooks(directory):
    """Check all incident management workbooks in a directory."""

    print("=" * 80)
    print("A.5.24-28 INCIDENT MANAGEMENT - BATCH SANITY CHECK")
    print("=" * 80)
    print(f"Directory: {directory}\n")

    # Find all workbooks
    patterns = [
        os.path.join(directory, "ISMS-IMP-A.5.24-28.S*.xlsx"),
        os.path.join(directory, "ISMS-IMP-A.5.24-28.S*_*.xlsx"),
    ]

    workbooks = []
    for pattern in patterns:
        workbooks.extend(glob(pattern))

    workbooks = sorted(set(workbooks))

    if not workbooks:
        print("[WARNING] No A.5.24-28 workbooks found in directory")
        return 1

    print(f"Found {len(workbooks)} workbook(s):\n")
    for wb_path in workbooks:
        print(f"  - {os.path.basename(wb_path)}")

    print("\n" + "=" * 80 + "\n")

    results = {}
    max_exit_code = 0

    for wb_path in workbooks:
        wb_type, definition = detect_workbook_type(wb_path)
        if definition:
            exit_code = check_workbook(wb_path, wb_type, definition)
            results[os.path.basename(wb_path)] = exit_code
            max_exit_code = max(max_exit_code, exit_code)
            print("\n")

    # Summary
    print("=" * 80)
    print("BATCH VALIDATION SUMMARY")
    print("=" * 80)

    for filename, code in results.items():
        status = "[PASS]" if code == 0 else "[WARNING]" if code == 1 else "[CRITICAL]"
        print(f"  {status} {filename}")

    print("\n" + "=" * 80)

    return max_exit_code


def main():
    if len(sys.argv) < 2:
        print("=" * 80)
        print("A.5.24-28 INCIDENT MANAGEMENT LIFECYCLE - SANITY CHECK")
        print("=" * 80)
        print("\nUsage:")
        print("  python3 sanity_check_a524_28.py <filename.xlsx>")
        print("  python3 sanity_check_a524_28.py --all <directory>")
        print("\nSupported Workbooks:")
        print("  S1: Framework Assessment (10 sheets)")
        print("  S2: Detection & Classification (9 sheets)")
        print("  S3: Response Capabilities (11 sheets)")
        print("  S4: Forensic Evidence (10 sheets)")
        print("  S5: Learning & Improvement (10 sheets)")
        print("\nExamples:")
        print("  python3 sanity_check_a524_28.py ISMS-IMP-A.5.24-28.S1_Framework_Assessment_20260131.xlsx")
        print("  python3 sanity_check_a524_28.py --all ../90_workbooks/")
        print("\nExit codes:")
        print("  0 = All checks passed")
        print("  1 = Warnings detected (workbook usable)")
        print("  2 = Critical errors detected (regenerate recommended)")
        print("\n" + "=" * 80)
        sys.exit(1)

    if sys.argv[1] == "--all":
        if len(sys.argv) < 3:
            directory = "../90_workbooks/"
        else:
            directory = sys.argv[2]
        exit_code = check_all_workbooks(directory)
    else:
        filename = sys.argv[1]
        exit_code = check_workbook(filename)

    sys.exit(exit_code)


if __name__ == "__main__":
    main()


# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (syntax validated, structure verified)
# QA_TOOL: Claude Code Deep Scan
# QA_NOTE: STANDARDIZATION - License header, logging, main() pattern applied
# =============================================================================
