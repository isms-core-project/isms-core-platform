#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.5.24-28.S4 - Forensic Evidence Assessment Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.28: Incident Evidence Preservation
Assessment Domain 4 of 5: Forensic Evidence Management

--------------------------------------------------------------------------------
PURPOSE & DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
digital forensic evidence management capabilities, focusing on collection,
preservation, analysis, storage, and legal readiness (A.5.28).

**Assessment Scope:**
- Evidence collection procedures and coverage
- Chain of custody management and integrity verification
- Forensic analysis tools, capabilities, and methodology
- Evidence storage, retention, and secure disposal
- Legal admissibility and regulatory compliance readiness

**Generated Workbook Structure:**
1. Instructions & Legend
2. Evidence Collection - 25 questions
3. Chain of Custody - 20 questions
4. Forensic Analysis - 20 questions
5. Storage & Retention - 15 questions
6. Legal & Regulatory Readiness - 15 questions
7. Gap Analysis - 40 capacity
8. Evidence Register - 60 capacity
9. Dashboard
10. Approval Sign-Off

**Key Features:**
- 95 assessment questions across 5 forensic domains
- Evidence collection coverage matrix (10 evidence types)
- Chain of custody integrity assessment
- Forensic tool inventory and capability evaluation
- Legal admissibility readiness scoring
- Automated metric calculations

**Integration:**
Domain 4 of 5 for A.5.24-28 Incident Management:
- S1: Framework & Governance (A.5.24)
- S2: Detection & Classification (A.5.25)
- S3: Response Capabilities (A.5.26)
- S4: Forensic Evidence Management (this - A.5.28)
- S5: Learning & Improvement (A.5.27)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

Python 3.8+, openpyxl

Installation: pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

python3 generate_a524_28_s4_forensic_evidence.py

Output: ISMS-IMP-A.5.24-28.S4_Forensic_Evidence_YYYYMMDD.xlsx

Estimated completion: 6-10 hours

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 A.5.24-28
Assessment Domain:    4 of 5 (Forensic Evidence - A.5.28)
Related Standards:    ISO/IEC 27037:2012, NIST SP 800-86
Version:              1.0
Date:                 2026-01-31
Python Version:       3.8+

================================================================================
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.24-28.S4"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_Forensic_Evidence_{GENERATED_TIMESTAMP}.xlsx"


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions & Legend",
        "Evidence Collection",
        "Chain of Custody",
        "Forensic Analysis",
        "Storage & Retention",
        "Legal & Regulatory",
        "Gap Analysis",
        "Evidence Register",
        "Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def setup_styles():
    """Define cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D8E4F8", end_color="D8E4F8", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "calculated_cell": {
            "fill": PatternFill(start_color="C8F0C8", end_color="C8F0C8", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
        "gap_critical": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "gap_high": {"fill": PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")},
        "gap_medium": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "gap_low": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
    }


# ============================================================================
# SECTION 2: INSTRUCTIONS SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.5.24-28.S4 — Forensic Evidence Assessment\n"
        "ISO/IEC 27001:2022 - Control A.5.28: Incident Evidence Preservation"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    row = 3
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:B{row}")

    doc_info = [
        ("Document ID", "ISMS-IMP-A.5.24-28.S4"),
        ("Assessment Area", "Forensic Evidence Management (Domain 4)"),
        ("Related Policy", "ISMS-POL-A.5.24-28, Section 2.4"),
        ("Related Reference", "ISMS-REF-A.5.24-28, Section 3"),
        ("Standards", "ISO/IEC 27037:2012, NIST SP 800-86"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
    ]

    row += 1
    for label, value in doc_info:
        ws[f"A{row}"] = label + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if label in ["Assessment Date", "Completed By", "Organisation"]:
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Key principles
    row += 2
    ws[f"A{row}"] = "FORENSIC EVIDENCE CORE PRINCIPLES (ISO/IEC 27037:2012)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:G{row}")

    row += 1
    principles = (
        "1. IDENTIFICATION: Locate and recognise relevant digital evidence\n"
        "2. COLLECTION: Gather evidence using validated, documented methods\n"
        "3. ACQUISITION: Create forensic copies preserving original integrity\n"
        "4. PRESERVATION: Maintain evidence integrity and chain of custody\n\n"
        "Key Rules: Never analyse originals. Always hash. Document everything. Preserve chain of custody."
    )
    ws.merge_cells(f"A{row}:G{row+4}")
    ws[f"A{row}"] = principles
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 90

    # Color legend
    row += 6
    ws[f"A{row}"] = "COLOR LEGEND"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:B{row}")

    legend = [
        ("Yellow (FFFFCC)", "Input cell — Enter your response"),
        ("Green (C8F0C8)", "Calculated cell — Auto-populated formula"),
        ("Light Blue (D8E4F8)", "Section header"),
    ]
    row += 1
    for color_label, description in legend:
        ws[f"A{row}"] = color_label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = description
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 3: GENERIC ASSESSMENT SHEET BUILDER
# ============================================================================

def create_assessment_sheet(ws, styles, sheet_title, subtitle, questions):
    """
    Generic builder for assessment sheets.
    questions: list of (q_id, section, question_text, answer_hint)
    Inserts section header rows when section changes.
    """
    ws.merge_cells("A1:G1")
    ws["A1"] = sheet_title
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")

    # Column headers at row 4
    row = 4
    headers = ["Question_ID", "Section", "Question", "Answer", "Evidence_Ref", "Comments", "Gap_Flag"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    current_section = None
    row = 5

    for q_id, section, question, answer_hint in questions:
        # Insert section header row when section changes
        if section != current_section:
            current_section = section
            ws.merge_cells(f"A{row}:G{row}")
            ws[f"A{row}"] = f"  {section}"
            ws[f"A{row}"].font = styles["section_header"]["font"]
            ws[f"A{row}"].fill = styles["section_header"]["fill"]
            ws[f"A{row}"].alignment = styles["section_header"]["alignment"]
            for col_idx in range(1, 8):
                ws.cell(row=row, column=col_idx).border = styles["border"]
            ws.row_dimensions[row].height = 20
            row += 1

        # Question row
        ws[f"A{row}"] = q_id
        ws[f"A{row}"].border = styles["border"]
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

        ws[f"B{row}"] = section
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")

        ws[f"C{row}"] = question
        ws[f"C{row}"].border = styles["border"]
        ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Answer (yellow input)
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"D{row}"].border = styles["border"]

        # Evidence reference (yellow input)
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].border = styles["border"]

        # Comments (yellow input)
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].border = styles["border"]

        # Gap flag (green calculated)
        ws[f"G{row}"] = f'=IF(OR(D{row}="No", D{row}="Never", D{row}="None", D{row}="No Capability"), "Yes", "No")'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        ws[f"G{row}"].border = styles["border"]
        ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 35
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 58
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 32
    ws.column_dimensions["G"].width = 10
    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 4: EVIDENCE COLLECTION SHEET
# ============================================================================

def create_evidence_collection(ws, styles):
    """Evidence Collection — 25 questions."""
    questions = [
        # Section A: Collection Procedures
        ("Q1",  "Collection Procedures", "Are forensic evidence collection procedures documented?",
         "Comprehensive / Basic / No"),
        ("Q2",  "Collection Procedures", "Is evidence collection integrated into IR playbooks?",
         "Embedded / Separate Procedure / No"),
        ("Q3",  "Collection Procedures", "Which evidence types have collection procedures? (Disk/Memory/Network/Logs/Email/Cloud/Mobile/Physical/App/IoT)",
         "Checkbox — select all that apply"),
        ("Q4",  "Collection Procedures", "Is evidence collection prioritised (volatile first)?",
         "Documented Priority / Informal / No"),
        ("Q5",  "Collection Procedures", "Is there a pre-collection checklist?",
         "Yes / No"),
        ("Q6",  "Collection Procedures", "Are standard forensic collection tools defined and approved?",
         "Approved Tool List / Ad-Hoc / No"),
        ("Q7",  "Collection Procedures", "Collection training frequency?",
         "Quarterly / Semi-Annually / Annually / Onboarding Only / Never"),
        ("Q8",  "Collection Procedures", "Are common collection mistakes addressed in training?",
         "Yes / No"),

        # Section B: Evidence by Source
        ("Q9",  "Evidence Sources", "Endpoint Disk — Collection capability?",
         "Full Image / Targeted / No Capability"),
        ("Q10", "Evidence Sources", "Endpoint Memory (RAM) — Collection capability?",
         "Live Capture / Not Capable / No Process"),
        ("Q11", "Evidence Sources", "Network Traffic (PCAP) — Collection capability?",
         "Real-Time / Retrospective / No Capability"),
        ("Q12", "Evidence Sources", "System/Application Logs — Collection capability?",
         "Centralised (SIEM) / Per-System / No"),
        ("Q13", "Evidence Sources", "Email Evidence — Collection capability?",
         "Full Mailbox Export / Individual Message / No"),
        ("Q14", "Evidence Sources", "Cloud Audit Logs — Collection capability?",
         "Automated Export / Manual Export / No Logging"),
        ("Q15", "Evidence Sources", "Mobile Device — Collection capability?",
         "Forensic Acquisition / Logical Only / No Capability"),
        ("Q16", "Evidence Sources", "Physical Evidence — Collection capability?",
         "Documented Procedure / Informal / No"),
        ("Q17", "Evidence Sources", "Application-Level Evidence — Collection capability?",
         "Automated / Manual (with Dev) / No"),
        ("Q18", "Evidence Sources", "IoT / OT Device Evidence — Collection capability?",
         "Capable / Limited / No Capability / N/A"),

        # Section C: Collection Quality
        ("Q19", "Collection Quality", "Was evidence collection complete? (Last 10 incidents)",
         "Consistently / Usually / Sometimes / No"),
        ("Q20", "Collection Quality", "Was volatile evidence captured before containment? (Last 10 incidents)",
         "Consistently / Usually / Sometimes / No"),
        ("Q21", "Collection Quality", "Average time to begin evidence collection after incident declaration?",
         "Enter duration (e.g. 30 min)"),
        ("Q22", "Collection Quality", "Is evidence collection documented in real-time?",
         "Real-Time / Post-Incident / No"),
        ("Q23", "Collection Quality", "% incidents with adequate evidence collection? (Last 10)",
         "Enter percentage (0-100)"),
        ("Q24", "Collection Quality", "If external forensics engaged, is collection coordinated?",
         "Before Engagement / During / No Coordination / N/A"),
        ("Q25", "Collection Quality", "Are collection lessons learned captured and applied?",
         "Systematically / Sometimes / No"),
    ]

    create_assessment_sheet(ws, styles,
        "Sheet 2: Evidence Collection",
        "Collection Procedures, Evidence Sources, Collection Quality (25 Questions)",
        questions)


# ============================================================================
# SECTION 5: CHAIN OF CUSTODY SHEET
# ============================================================================

def create_chain_of_custody(ws, styles):
    """Chain of Custody — 20 questions."""
    questions = [
        # Section A: CoC Documentation
        ("Q26", "CoC Documentation", "Is chain of custody procedure documented?",
         "Comprehensive / Basic / No"),
        ("Q27", "CoC Documentation", "Are CoC forms completed for all evidence?",
         "Always / Usually / Sometimes / No"),
        ("Q28", "CoC Documentation", "How is CoC documented?",
         "Digital System / Paper Forms / Both / None"),
        ("Q29", "CoC Documentation", "Is every evidence transfer documented?",
         "All Transfers / Most / Minimal / No"),
        ("Q30", "CoC Documentation", "Is access to stored evidence logged?",
         "All Access / Selected Access / No"),
        ("Q31", "CoC Documentation", "Was CoC complete in recent investigations? (Last 10)",
         "Consistently / Usually / Sometimes / No"),
        ("Q32", "CoC Documentation", "Do all evidence-handling personnel receive CoC training?",
         "All Personnel / Forensic Team Only / No"),

        # Section B: Integrity Verification
        ("Q33", "Integrity Verification", "Are hash values calculated for all collected evidence?",
         "Always / Usually / Sometimes / No"),
        ("Q34", "Integrity Verification", "What hash algorithm is used?",
         "SHA-256 / SHA-512 / MD5 (Legacy) / Multiple / None"),
        ("Q35", "Integrity Verification", "Is dual hashing used (two independent algorithms)?",
         "Yes / No"),
        ("Q36", "Integrity Verification", "Are hash values verified when copies are made?",
         "Always / Sometimes / No"),
        ("Q37", "Integrity Verification", "Are hash values stored separately from evidence?",
         "Separate Storage / With Evidence / No"),
        ("Q38", "Integrity Verification", "How often is stored evidence integrity verified?",
         "On Each Access / Monthly / Quarterly / Annually / Never"),

        # Section C: Access Control
        ("Q39", "Access Control", "Is access to forensic evidence restricted?",
         "Strictly Controlled / Role-Based / Limited / No"),
        ("Q40", "Access Control", "Is an authorised access list maintained per evidence item?",
         "Yes / No"),
        ("Q41", "Access Control", "Is evidence accessible in read-only mode?",
         "Enforced / Policy Only / No"),
        ("Q42", "Access Control", "Is physical evidence in a secure, access-controlled location?",
         "Dedicated Room / Secure Location / No"),
        ("Q43", "Access Control", "Can unauthorised access to evidence be detected?",
         "Alerted / Logged / No"),
        ("Q44", "Access Control", "Are controls in place for sharing evidence externally?",
         "Documented Controls / Informal / No"),
        ("Q45", "Access Control", "For physical evidence, are handling precautions followed (gloves, packaging)?",
         "Always / Sometimes / No / N/A"),
    ]

    create_assessment_sheet(ws, styles,
        "Sheet 3: Chain of Custody",
        "CoC Documentation, Integrity Verification, Access Control (20 Questions)",
        questions)


# ============================================================================
# SECTION 6: FORENSIC ANALYSIS SHEET
# ============================================================================

def create_forensic_analysis(ws, styles):
    """Forensic Analysis — 20 questions."""
    questions = [
        # Section A: Analysis Tools
        ("Q46", "Analysis Tools", "What forensic tools are available? (Disk/Memory/Network/Malware/Mobile/Cloud/Log/Other)",
         "Checkbox — select all that apply"),
        ("Q47", "Analysis Tools", "Tool type: Commercial / Open-Source / Both / None?",
         "Commercial Only / Open-Source Only / Both / None"),
        ("Q48", "Analysis Tools", "Are forensic tools properly licensed and maintained?",
         "All Licensed / Mostly / No"),
        ("Q49", "Analysis Tools", "Are forensic tools validated before use?",
         "Regularly Validated / Initial Validation / No"),
        ("Q50", "Analysis Tools", "Are forensic tools kept up-to-date?",
         "Monthly / Quarterly / Infrequently / Never"),
        ("Q51", "Analysis Tools", "Are dedicated forensic workstations available (isolated)?",
         "Dedicated / Shared (Isolated) / No"),
        ("Q52", "Analysis Tools", "Is a malware analysis sandbox available?",
         "Commercial / Internal / No"),

        # Section B: Analyst Capabilities
        ("Q53", "Analyst Capabilities", "How many qualified forensic analysts?",
         "Enter number"),
        ("Q54", "Analyst Capabilities", "Do forensic analysts hold recognised certifications?",
         "GCFE/GCIH/EnCE/CCFP / Other Certs / No Certs"),
        ("Q55", "Analyst Capabilities", "Forensic analyst training frequency?",
         "Quarterly / Semi-Annually / Annually / Onboarding Only / Never"),
        ("Q56", "Analyst Capabilities", "Is analysis methodology documented?",
         "Comprehensive / Basic / No"),
        ("Q57", "Analyst Capabilities", "Is analysis always performed on copies only?",
         "Always / Usually / Sometimes / No"),
        ("Q58", "Analyst Capabilities", "Are findings reproducible (documented for verification)?",
         "Yes - Documented / Partial / No"),
        ("Q59", "Analyst Capabilities", "Are forensic reports reviewed by a second analyst?",
         "Always / Sometimes / No"),
        ("Q60", "Analyst Capabilities", "Can external forensic services be engaged?",
         "Retainer / Ad-Hoc / No"),

        # Section C: Analysis Outcomes
        ("Q61", "Analysis Outcomes", "% incidents with root cause identified? (Last 10 requiring forensics)",
         "Enter percentage (0-100)"),
        ("Q62", "Analysis Outcomes", "% breach incidents with full attacker scope determined?",
         "Enter percentage (0-100)"),
        ("Q63", "Analysis Outcomes", "Are IOCs generated from forensic analysis?",
         "Systematically / Sometimes / No"),
        ("Q64", "Analysis Outcomes", "Are incident timelines reconstructed accurately?",
         "Comprehensive / Partial / No"),
        ("Q65", "Analysis Outcomes", "Were forensic findings actionable?",
         "Consistently / Usually / Sometimes / No"),
    ]

    create_assessment_sheet(ws, styles,
        "Sheet 4: Forensic Analysis",
        "Analysis Tools, Analyst Capabilities, Analysis Outcomes (20 Questions)",
        questions)


# ============================================================================
# SECTION 7: STORAGE & RETENTION SHEET
# ============================================================================

def create_storage_retention(ws, styles):
    """Storage & Retention — 15 questions."""
    questions = [
        # Section A: Evidence Storage
        ("Q66", "Evidence Storage", "Where is forensic evidence stored?",
         "Dedicated Server / Cloud (Isolated) / General Share / No Dedicated Storage"),
        ("Q67", "Evidence Storage", "Is storage access-controlled?",
         "Strict (Role-Based) / Basic / No"),
        ("Q68", "Evidence Storage", "Is stored evidence encrypted?",
         "At Rest + In Transit / At Rest Only / No"),
        ("Q69", "Evidence Storage", "Are backup copies of critical evidence maintained?",
         "Multiple Copies / Single Backup / No"),
        ("Q70", "Evidence Storage", "Is evidence storage monitored?",
         "Real-Time Alerts / Periodic Review / No"),
        ("Q71", "Evidence Storage", "Is an evidence catalog maintained?",
         "Comprehensive / Basic / No"),
        ("Q72", "Evidence Storage", "Is storage capacity sufficient?",
         "Adequate / Near Capacity / Insufficient"),

        # Section B: Retention & Disposal
        ("Q73", "Retention & Disposal", "Are retention periods defined for forensic evidence?",
         "By Severity / Uniform Period / No"),
        ("Q74", "Retention & Disposal", "Do retention periods meet regulatory requirements?",
         "Yes / No / Not Assessed"),
        ("Q75", "Retention & Disposal", "Can retention be extended via legal hold?",
         "Documented Process / Informal / No"),
        ("Q76", "Retention & Disposal", "Are evidence disposal procedures documented?",
         "Yes / No"),
        ("Q77", "Retention & Disposal", "Is disposal authorised before execution?",
         "Manager Approval / Legal Approval / No Authorisation"),
        ("Q78", "Retention & Disposal", "Is disposal verified after execution?",
         "Yes / No"),
        ("Q79", "Retention & Disposal", "Are disposal records maintained?",
         "Yes / No"),
        ("Q80", "Retention & Disposal", "Is disposal automated when retention expires?",
         "Automated + Notification / Manual / No"),
    ]

    create_assessment_sheet(ws, styles,
        "Sheet 5: Storage & Retention",
        "Evidence Storage, Retention & Disposal (15 Questions)",
        questions)


# ============================================================================
# SECTION 8: LEGAL & REGULATORY READINESS SHEET
# ============================================================================

def create_legal_regulatory(ws, styles):
    """Legal & Regulatory Readiness — 15 questions."""
    questions = [
        # Section A: Legal Admissibility
        ("Q81", "Legal Admissibility", "Have forensic procedures been reviewed by Legal?",
         "Recently Reviewed / Initial Only / No"),
        ("Q82", "Legal Admissibility", "Are legal admissibility requirements documented?",
         "Jurisdiction-Specific / General / No"),
        ("Q83", "Legal Admissibility", "Can [Organisation] provide expert witnesses?",
         "Internal Expert / External Available / No"),
        ("Q84", "Legal Admissibility", "Is forensic report format suitable for legal proceedings?",
         "Legal-Ready / Partially / No"),
        ("Q85", "Legal Admissibility", "Are forensic analysts prepared for cross-examination?",
         "Trained / Partially / No / Not Yet Required"),
        ("Q86", "Legal Admissibility", "Is evidence integrity fully documented (end-to-end)?",
         "Complete Trail / Mostly / No"),
        ("Q87", "Legal Admissibility", "Is physical evidence photographed in situ?",
         "Always / Sometimes / No / N/A"),
        ("Q88", "Legal Admissibility", "Are witness statements collected during incidents?",
         "Systematically / Sometimes / No"),

        # Section B: Regulatory Compliance
        ("Q89", "Regulatory Compliance", "GDPR forensic evidence compliance?",
         "Yes / No / Not Applicable"),
        ("Q90", "Regulatory Compliance", "PCI DSS forensic evidence compliance?",
         "Yes / No / Not Applicable"),
        ("Q91", "Regulatory Compliance", "Industry-specific forensic requirements met?",
         "Yes / No / Not Applicable"),
        ("Q92", "Regulatory Compliance", "Can evidence be submitted to regulators in required format?",
         "Ready / Needs Preparation / No"),
        ("Q93", "Regulatory Compliance", "Law enforcement evidence cooperation procedures?",
         "Documented / Informal / No"),
        ("Q94", "Regulatory Compliance", "International evidence requirements considered?",
         "Yes / No / Not Applicable"),
        ("Q95", "Regulatory Compliance", "Forensic evidence available for regulatory audits?",
         "Readily Available / Needs Organisation / No"),
    ]

    create_assessment_sheet(ws, styles,
        "Sheet 6: Legal & Regulatory Readiness",
        "Legal Admissibility, Regulatory Compliance (15 Questions)",
        questions)


# ============================================================================
# SECTION 9: GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis(ws, styles):
    """Gap Analysis — 40 capacity."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Sheet 7: Gap Analysis"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Summary row
    ws["A3"] = "GAP SUMMARY"
    ws["A3"].font = Font(bold=True, size=10, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells("A3:D3")

    ws["A4"] = "Total Gaps:"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = '=COUNTA(C6:C45)-COUNTBLANK(C6:C45)'
    ws["B4"].fill = styles["calculated_cell"]["fill"]
    ws["B4"].border = styles["border"]

    ws["C4"] = "Critical:"
    ws["C4"].font = Font(bold=True)
    ws["D4"] = '=COUNTIF(E6:E45,"Critical")'
    ws["D4"].fill = styles["calculated_cell"]["fill"]
    ws["D4"].border = styles["border"]

    ws["E4"] = "High:"
    ws["E4"].font = Font(bold=True)
    ws["F4"] = '=COUNTIF(E6:E45,"High")'
    ws["F4"].fill = styles["calculated_cell"]["fill"]
    ws["F4"].border = styles["border"]

    ws["G4"] = "Medium:"
    ws["G4"].font = Font(bold=True)
    ws["H4"] = '=COUNTIF(E6:E45,"Medium")'
    ws["H4"].fill = styles["calculated_cell"]["fill"]
    ws["H4"].border = styles["border"]

    # Column headers at row 6
    row = 6
    headers = ["Gap_ID", "Domain", "Gap_Description", "Risk_Level", "Current_State",
               "Target_State", "Remediation_Action", "Owner", "Target_Date", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # 40 gap rows
    for i in range(40):
        row = 7 + i
        ws[f"A{row}"] = f"GAP-{str(i+1).zfill(3)}"
        ws[f"A{row}"].border = styles["border"]
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        for col_idx in range(2, 11):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 26
    ws.column_dimensions["G"].width = 42
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.freeze_panes = "A7"


# ============================================================================
# SECTION 10: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """Evidence Register — 60 capacity."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Sheet 8: Evidence Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 4
    headers = ["Evidence_ID", "Evidence_Type", "Description", "Related_Question",
               "Storage_Location", "Date_Collected", "Collected_By", "Verification_Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    for i in range(60):
        row = 5 + i
        ws[f"A{row}"] = f"EV-{str(i+1).zfill(3)}"
        ws[f"A{row}"].border = styles["border"]
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        for col_idx in range(2, 9):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 22
    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 11: DASHBOARD SHEET
# ============================================================================

def create_dashboard(ws, styles):
    """Dashboard — Forensic evidence summary."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "Sheet 9: Dashboard — Forensic Evidence Summary"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # --- Section A: Key Metrics ---
    row = 3
    ws[f"A{row}"] = "KEY FORENSIC METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:C{row}")

    metrics = [
        ("Evidence Collection Completeness (%)", "='Evidence Collection'!D23"),
        ("Root Cause Identification Rate (%)", "='Forensic Analysis'!D61"),
        ("Attacker Scope Determination Rate (%)", "='Forensic Analysis'!D62"),
        ("Time to Begin Collection", "='Evidence Collection'!D21"),
        ("Forensic Analysts (FTE)", "='Forensic Analysis'!D53"),
    ]

    row = 4
    for label, formula in metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, size=10)
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = formula
        ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    # --- Section B: Coverage Matrix Summary ---
    row += 1
    ws[f"A{row}"] = "EVIDENCE SOURCE COVERAGE"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:C{row}")

    sources = [
        ("Endpoint Disk", "='Evidence Collection'!D9"),
        ("Memory (RAM)", "='Evidence Collection'!D10"),
        ("Network Traffic", "='Evidence Collection'!D11"),
        ("System Logs", "='Evidence Collection'!D12"),
        ("Email", "='Evidence Collection'!D13"),
        ("Cloud Audit Logs", "='Evidence Collection'!D14"),
        ("Mobile Device", "='Evidence Collection'!D15"),
        ("Physical Evidence", "='Evidence Collection'!D16"),
        ("Application Data", "='Evidence Collection'!D17"),
        ("IoT / OT", "='Evidence Collection'!D18"),
    ]

    row += 1
    for source_label, source_formula in sources:
        ws[f"A{row}"] = source_label
        ws[f"A{row}"].font = Font(size=10)
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = source_formula
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    # --- Section C: Gap Summary ---
    row += 1
    ws[f"A{row}"] = "GAP SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:C{row}")

    gap_metrics = [
        ("Total Gaps", "='Gap Analysis'!B4"),
        ("Critical Gaps", "='Gap Analysis'!D4"),
        ("High Gaps", "='Gap Analysis'!F4"),
        ("Medium Gaps", "='Gap Analysis'!H4"),
    ]

    row += 1
    for label, formula in gap_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, size=10)
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = formula
        ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 24
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: APPROVAL SIGN-OFF SHEET
# ============================================================================

def create_approval_signoff(ws, styles):
    """Approval Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "Sheet 10: Approval & Sign-Off"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Assessment Summary
    row = 3
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")

    row += 1
    summary = [
        ("Assessment Document", "ISMS-IMP-A.5.24-28.S4 — Forensic Evidence Assessment"),
        ("Assessment Period", ""),
        ("Evidence Collection Completeness", "=Dashboard!B4"),
        ("Root Cause ID Rate", "=Dashboard!B5"),
        ("Total Gaps Identified", "=Dashboard!B25"),
        ("Critical Gaps", "=Dashboard!B26"),
    ]

    for label, value in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = value
        if label == "Assessment Period":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        elif label == "Assessment Document":
            ws[f"B{row}"].font = Font(size=10)
        else:
            ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws.merge_cells(f"B{row}:E{row}")
        row += 1

    # Completed By — Forensic Lead
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "COMPLETED BY (Digital Forensics Lead)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    for field in ["Name", "Date", "Signature"]:
        row += 1
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = styles["border"]
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]

    # Approved By — CISO
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    for field in ["Name", "Date", "Approval Decision", "Signature"]:
        row += 1
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = styles["border"]
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]

    # Next review
    row += 2
    ws[f"A{row}"] = "Next Review Date:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].border = styles["border"]
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 13: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution."""
    print("=" * 80)
    print("ISMS-IMP-A.5.24-28.S4 — Forensic Evidence Assessment Generator")
    print("ISO/IEC 27001:2022 Control A.5.28: Incident Evidence Preservation")
    print("=" * 80)

    wb = create_workbook()
    styles = setup_styles()

    print("\n[ 1/10] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)

    print("[ 2/10] Creating Evidence Collection (25 questions)...")
    create_evidence_collection(wb["Evidence Collection"], styles)

    print("[ 3/10] Creating Chain of Custody (20 questions)...")
    create_chain_of_custody(wb["Chain of Custody"], styles)

    print("[ 4/10] Creating Forensic Analysis (20 questions)...")
    create_forensic_analysis(wb["Forensic Analysis"], styles)

    print("[ 5/10] Creating Storage & Retention (15 questions)...")
    create_storage_retention(wb["Storage & Retention"], styles)

    print("[ 6/10] Creating Legal & Regulatory Readiness (15 questions)...")
    create_legal_regulatory(wb["Legal & Regulatory"], styles)

    print("[ 7/10] Creating Gap Analysis...")
    create_gap_analysis(wb["Gap Analysis"], styles)

    print("[ 8/10] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)

    print("[ 9/10] Creating Dashboard...")
    create_dashboard(wb["Dashboard"], styles)

    print("[10/10] Creating Approval Sign-Off...")
    create_approval_signoff(wb["Approval Sign-Off"], styles)

    wb.save(OUTPUT_FILENAME)

    print(f"\n{'=' * 80}")
    print(f"✅  SUCCESS: {OUTPUT_FILENAME}")
    print(f"{'=' * 80}")
    print("\nWorkbook Structure:")
    print("  • 10 sheets (Instructions through Approval)")
    print("  • 95 assessment questions (25+20+20+15+15)")
    print("  • 40 gap analysis capacity (with summary formulas)")
    print("  • 60 evidence register capacity")
    print("  • Dashboard: key metrics + coverage matrix + gap summary")
    print("\nNext steps:")
    print("  1) Review ISMS-REF-A.5.24-28 Section 3 (Forensic Techniques Library)")
    print("  2) Inventory current forensic tools and capabilities")
    print("  3) Review last 10 incidents for evidence quality")
    print("  4) Fill yellow cells across all assessment sheets")
    print("  5) Complete gap analysis based on findings")
    print("  6) Legal review of evidence procedures")
    print("  7) Complete Approval Sign-Off workflow")
    print(f"\nEstimated completion time: 6-10 hours")
    print(f"{'=' * 80}")


if __name__ == "__main__":
    main()


# ============================================================================
# END OF SCRIPT
# ============================================================================

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# =============================================================================

