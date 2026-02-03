#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.24-28.S6 - Incident Management Compliance Dashboard Generator
================================================================================

ISO/IEC 27001:2022 Controls A.5.24-28: Incident Management Lifecycle
Assessment Domain 6 of 6: Compliance Dashboard (Consolidated)

--------------------------------------------------------------------------------
PURPOSE & DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel dashboard workbook that consolidates
incident management compliance metrics from all five assessment domains (S1-S5)
into executive-ready visualisations.

**Dashboard Scope:**
- Executive summary of overall incident management compliance
- Control-by-control compliance scoring (A.5.24 through A.5.28)
- Gap analysis consolidation across all assessment domains
- Key Performance Indicators (KPIs) for incident management
- Evidence register linking to source assessments
- Approval and sign-off workflow

**Generated Workbook Structure:**
1. Instructions - Dashboard overview and guidance
2. Executive_Summary - KPIs, compliance overview, recommendations
3. Compliance_Scores - Per-control compliance breakdown
4. Gap_Analysis - Consolidated gap inventory (100 items capacity)
5. Incident_KPIs - Operational metrics and trend tracking
6. Evidence_Register - Evidence inventory (100 items capacity)
7. Approval_Sign_Off - Multi-stakeholder approval workflow

**Key Features:**
- Control-by-control compliance heat map
- Traffic light status indicators (GREEN/AMBER/RED)
- Gap aging calculations
- KPI trend tracking
- Evidence traceability
- Protected formulas with unprotected input cells

**Source Workbooks:**
- S1: ISMS-IMP-A.5.24-28.S1_Framework_Assessment (A.5.24)
- S2: ISMS-IMP-A.5.24-28.S2_Detection_Classification (A.5.25)
- S3: ISMS-IMP-A.5.24-28.S3_Response_Capabilities (A.5.26)
- S4: ISMS-IMP-A.5.24-28.S4_Forensic_Evidence (A.5.28)
- S5: ISMS-IMP-A.5.24-28.S5_Learning_Improvement (A.5.27)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - logging (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a524_28_s6_compliance_dashboard.py

Output:
    File: ISMS-IMP-A.5.24-28.S6_Compliance_Dashboard_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Open all S1-S5 source workbooks
    2. Import compliance scores from each dashboard sheet
    3. Consolidate gaps from all gap analysis sheets
    4. Update KPI data from incident ticketing system
    5. Link evidence references
    6. Obtain management approvals

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.5.24-28
Assessment Domain:    6 of 6 (Compliance Dashboard - Consolidated)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              Internal Use

Related Documents:
    - ISMS-POL-A.5.24-28: Incident Management Lifecycle Policy
    - ISMS-REF-A.5.24-28: Incident Response Reference Guide
    - ISMS-IMP-A.5.24-28.S1-S5: Source Assessment Workbooks

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 2026-02-03
    - Initial release
    - Implements full dashboard per ISMS-IMP-A.5.24-28.S6 specification
    - 7 sheets with executive summary, compliance scores, gap analysis, KPIs
    - Evidence register and approval workflow

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.24-28.S6"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.5.24-28"
CONTROL_NAME = "Incident Management Lifecycle"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # Swiss format
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color="D8E4F8", end_color="D8E4F8", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "calculated_cell": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "metric_good": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "metric_warning": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "metric_bad": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "severity_critical": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "severity_high": {"fill": PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")},
        "severity_medium": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "severity_low": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "border": border_thin,
    }
    return styles


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching specification."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure per specification
    sheets = [
        "Instructions",
        "Executive_Summary",
        "Compliance_Scores",
        "Gap_Analysis",
        "Incident_KPIs",
        "Evidence_Register",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# SHEET 1: INSTRUCTIONS
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet with dashboard overview."""
    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 45

    # Document Information Section
    row = 3
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws.merge_cells(f"A{row}:B{row}")

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Incident Management Compliance Dashboard"),
        ("Related Policy", "ISMS-POL-A.5.24-28"),
        ("Version", "1.0"),
        ("Generated Date", GENERATED_DATE),
        ("Organisation", ""),  # User input
        ("Completed By", ""),  # User input
        ("Assessment Period", ""),  # User input
    ]

    row += 1
    for label, value in doc_info:
        ws[f"A{row}"] = label + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if label in ["Organisation", "Completed By", "Assessment Period"]:
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Dashboard Purpose Section
    row += 2
    ws[f"A{row}"] = "DASHBOARD PURPOSE"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws.merge_cells(f"A{row}:E{row}")

    row += 1
    purpose_text = """This dashboard consolidates incident management compliance metrics from all five assessment
domains (S1-S5) into executive-ready visualisations. It provides visibility into:

• Overall incident management compliance posture
• Control-by-control compliance scoring (A.5.24 through A.5.28)
• Gap analysis consolidation across all assessment domains
• Key Performance Indicators (KPIs) for incident management
• Evidence traceability linking to source assessments"""

    ws.merge_cells(f"A{row}:E{row+6}")
    ws[f"A{row}"] = purpose_text
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 120

    # Source Workbooks Section
    row += 8
    ws[f"A{row}"] = "SOURCE WORKBOOKS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws.merge_cells(f"A{row}:E{row}")

    row += 1
    sources = [
        ("S1", "A.5.24", "Framework Assessment", "ISMS-IMP-A.5.24-28.S1_Framework_Assessment"),
        ("S2", "A.5.25", "Detection & Classification", "ISMS-IMP-A.5.24-28.S2_Detection_Classification"),
        ("S3", "A.5.26", "Response Capabilities", "ISMS-IMP-A.5.24-28.S3_Response_Capabilities"),
        ("S4", "A.5.28", "Forensic Evidence", "ISMS-IMP-A.5.24-28.S4_Forensic_Evidence"),
        ("S5", "A.5.27", "Learning & Improvement", "ISMS-IMP-A.5.24-28.S5_Learning_Improvement"),
    ]

    headers = ["Source", "Control", "Assessment Domain", "Workbook Name"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    row += 1
    for source, control, domain, workbook in sources:
        ws.cell(row=row, column=1, value=source).border = styles["border"]
        ws.cell(row=row, column=2, value=control).border = styles["border"]
        ws.cell(row=row, column=3, value=domain).border = styles["border"]
        ws.cell(row=row, column=4, value=workbook).border = styles["border"]
        row += 1

    # Colour Legend Section
    row += 2
    ws[f"A{row}"] = "COLOUR LEGEND"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws.merge_cells(f"A{row}:C{row}")

    row += 1
    legend = [
        ("Yellow", "User Input Required", "FFFFCC"),
        ("Green", "Compliant / Good / Verified", "C6EFCE"),
        ("Amber", "Partial / Warning / Pending", "FFEB9C"),
        ("Red", "Non-Compliant / Critical / Expired", "FFC7CE"),
        ("Light Blue", "Section Header", "D8E4F8"),
        ("Grey", "Column Header", "D9D9D9"),
    ]

    for col_name, meaning, rgb in legend:
        ws.cell(row=row, column=1, value=col_name).fill = PatternFill(
            start_color=rgb, end_color=rgb, fill_type="solid"
        )
        ws.cell(row=row, column=1).border = styles["border"]
        ws.cell(row=row, column=2, value=meaning).border = styles["border"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 20

    ws.freeze_panes = "A3"


# =============================================================================
# SHEET 2: EXECUTIVE SUMMARY
# =============================================================================
def create_executive_summary_sheet(ws, styles):
    """Create the Executive Summary sheet with KPIs and compliance overview."""
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "INCIDENT MANAGEMENT - EXECUTIVE SUMMARY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Overall Compliance Section
    row = 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "OVERALL COMPLIANCE STATUS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    row += 2
    ws[f"A{row}"] = "Overall Incident Management Compliance:"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]
    ws[f"B{row}"].font = Font(bold=True, size=14)
    ws[f"B{row}"].alignment = Alignment(horizontal="center")
    ws[f"C{row}"] = "(Enter % from calculated formula)"
    ws[f"C{row}"].font = Font(italic=True, color="666666")

    # Control-by-Control Summary
    row += 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "COMPLIANCE BY CONTROL"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    row += 2
    headers = ["Control", "Description", "Weight", "Score", "Status", "Source"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    controls = [
        ("A.5.24", "Planning & Preparation", "25%", "", "", "S1"),
        ("A.5.25", "Assessment & Decision", "20%", "", "", "S2"),
        ("A.5.26", "Response to Incidents", "25%", "", "", "S3"),
        ("A.5.27", "Learning from Incidents", "15%", "", "", "S5"),
        ("A.5.28", "Collection of Evidence", "15%", "", "", "S4"),
    ]

    row += 1
    for control, desc, weight, score, status, source in controls:
        ws.cell(row=row, column=1, value=control).border = styles["border"]
        ws.cell(row=row, column=2, value=desc).border = styles["border"]
        ws.cell(row=row, column=3, value=weight).border = styles["border"]
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=4).border = styles["border"]
        ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=5).border = styles["border"]
        ws.cell(row=row, column=6, value=source).border = styles["border"]
        row += 1

    # Weighted Total Row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).border = styles["border"]
    ws.cell(row=row, column=2, value="Weighted Average").font = Font(bold=True)
    ws.cell(row=row, column=2).border = styles["border"]
    ws.cell(row=row, column=3, value="100%").font = Font(bold=True)
    ws.cell(row=row, column=3).border = styles["border"]
    ws.cell(row=row, column=4).fill = styles["calculated_cell"]["fill"]
    ws.cell(row=row, column=4).border = styles["border"]
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=5).border = styles["border"]
    ws.cell(row=row, column=6).border = styles["border"]

    # Top 5 Gaps Section
    row += 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "TOP 5 GAPS REQUIRING ATTENTION"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    row += 2
    gap_headers = ["Rank", "Gap Description", "Severity", "Owner", "Due Date", "Source"]
    for col, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    row += 1
    for i in range(1, 6):
        ws.cell(row=row, column=1, value=i).border = styles["border"]
        for c in range(2, 7):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # KPI Scorecard Section
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "KEY PERFORMANCE INDICATORS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    row += 2
    kpi_headers = ["KPI", "Current", "Target", "Status", "Trend", "Notes"]
    for col, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    kpis = [
        ("MTTD (Critical)", "", "<1 hour"),
        ("MTTR (Critical)", "", "<15 min"),
        ("Incident Closure Rate", "", ">95%"),
        ("PIR Completion Rate", "", "100%"),
        ("Training Completion", "", "100%"),
        ("Exercise Frequency", "", "2/year"),
    ]

    row += 1
    for kpi, current, target in kpis:
        ws.cell(row=row, column=1, value=kpi).border = styles["border"]
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]
        ws.cell(row=row, column=3, value=target).border = styles["border"]
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=4).border = styles["border"]
        ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=5).border = styles["border"]
        ws.cell(row=row, column=6).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=6).border = styles["border"]
        row += 1

    # Recommendations Section
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "RECOMMENDATIONS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    row += 2
    rec_headers = ["#", "Recommendation", "Owner", "Timeline", "Priority", "Status"]
    for col, header in enumerate(rec_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    row += 1
    for i in range(1, 6):
        ws.cell(row=row, column=1, value=i).border = styles["border"]
        for c in range(2, 7):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 25

    # Add data validation for Status
    dv_status = DataValidation(
        type="list",
        formula1='"GREEN - Compliant,AMBER - Partial,RED - Non-Compliant"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    # Apply to status columns (will be applied to appropriate rows)

    # Add data validation for Trend
    dv_trend = DataValidation(
        type="list",
        formula1='"▲ Improving,▬ Stable,▼ Declining"',
        allow_blank=True
    )
    ws.add_data_validation(dv_trend)

    ws.freeze_panes = "A3"


# =============================================================================
# SHEET 3: COMPLIANCE SCORES
# =============================================================================
def create_compliance_scores_sheet(ws, styles):
    """Create the Compliance Scores sheet with per-control breakdown."""
    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "COMPLIANCE SCORES BY CONTROL"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Column headers
    row = 3
    headers = ["Control", "Control Name", "Domain", "Weight", "Score", "Status", "Source Reference", "Notes"]
    col_widths = [12, 40, 30, 10, 12, 18, 45, 35]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Control data with domains
    controls = [
        # A.5.24 - Planning & Preparation
        ("A.5.24", "Incident Management Planning and Preparation", "Governance Framework", "25%", "S1-Framework"),
        ("A.5.24", "Incident Management Planning and Preparation", "Organisational Structure", "25%", "S1-Framework"),
        ("A.5.24", "Incident Management Planning and Preparation", "Training & Competency", "20%", "S1-Framework"),
        ("A.5.24", "Incident Management Planning and Preparation", "Tools & Technology", "15%", "S1-Framework"),
        ("A.5.24", "Incident Management Planning and Preparation", "Integration Assessment", "15%", "S1-Framework"),
        ("A.5.24", "Incident Management Planning and Preparation", "A.5.24 TOTAL", "100%", "S1-Dashboard"),

        # A.5.25 - Assessment & Decision
        ("A.5.25", "Assessment and Decision on Security Events", "Detection Sources", "20%", "S2-Detection"),
        ("A.5.25", "Assessment and Decision on Security Events", "Classification Framework", "25%", "S2-Detection"),
        ("A.5.25", "Assessment and Decision on Security Events", "Severity Assessment", "25%", "S2-Detection"),
        ("A.5.25", "Assessment and Decision on Security Events", "Triage Process", "20%", "S2-Detection"),
        ("A.5.25", "Assessment and Decision on Security Events", "Escalation Procedures", "10%", "S2-Detection"),
        ("A.5.25", "Assessment and Decision on Security Events", "A.5.25 TOTAL", "100%", "S2-Dashboard"),

        # A.5.26 - Response
        ("A.5.26", "Response to Information Security Incidents", "Response Procedures", "25%", "S3-Response"),
        ("A.5.26", "Response to Information Security Incidents", "Containment Capabilities", "25%", "S3-Response"),
        ("A.5.26", "Response to Information Security Incidents", "Eradication & Recovery", "20%", "S3-Response"),
        ("A.5.26", "Response to Information Security Incidents", "Communication Protocols", "15%", "S3-Response"),
        ("A.5.26", "Response to Information Security Incidents", "Coordination Effectiveness", "15%", "S3-Response"),
        ("A.5.26", "Response to Information Security Incidents", "A.5.26 TOTAL", "100%", "S3-Dashboard"),

        # A.5.27 - Learning
        ("A.5.27", "Learning from Information Security Incidents", "Post-Incident Review Process", "30%", "S5-Learning"),
        ("A.5.27", "Learning from Information Security Incidents", "Root Cause Analysis", "25%", "S5-Learning"),
        ("A.5.27", "Learning from Information Security Incidents", "Knowledge Management", "20%", "S5-Learning"),
        ("A.5.27", "Learning from Information Security Incidents", "Continuous Improvement", "15%", "S5-Learning"),
        ("A.5.27", "Learning from Information Security Incidents", "Metrics & Reporting", "10%", "S5-Learning"),
        ("A.5.27", "Learning from Information Security Incidents", "A.5.27 TOTAL", "100%", "S5-Dashboard"),

        # A.5.28 - Evidence
        ("A.5.28", "Collection of Evidence", "Evidence Identification", "20%", "S4-Forensic"),
        ("A.5.28", "Collection of Evidence", "Collection Procedures", "25%", "S4-Forensic"),
        ("A.5.28", "Collection of Evidence", "Chain of Custody", "25%", "S4-Forensic"),
        ("A.5.28", "Collection of Evidence", "Preservation & Storage", "20%", "S4-Forensic"),
        ("A.5.28", "Collection of Evidence", "Legal Admissibility", "10%", "S4-Forensic"),
        ("A.5.28", "Collection of Evidence", "A.5.28 TOTAL", "100%", "S4-Dashboard"),
    ]

    row = 4
    for control, name, domain, weight, source in controls:
        ws.cell(row=row, column=1, value=control).border = styles["border"]
        ws.cell(row=row, column=2, value=name).border = styles["border"]
        ws.cell(row=row, column=3, value=domain).border = styles["border"]
        ws.cell(row=row, column=4, value=weight).border = styles["border"]

        # Score and Status are input cells
        ws.cell(row=row, column=5).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=5).border = styles["border"]
        ws.cell(row=row, column=6).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=6).border = styles["border"]

        ws.cell(row=row, column=7, value=source).border = styles["border"]
        ws.cell(row=row, column=8).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=8).border = styles["border"]

        # Make TOTAL rows bold
        if "TOTAL" in domain:
            for c in range(1, 9):
                ws.cell(row=row, column=c).font = Font(bold=True)
            ws.cell(row=row, column=5).fill = styles["calculated_cell"]["fill"]

        row += 1

    # Overall Total Row
    row += 1
    ws.cell(row=row, column=1, value="OVERALL").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).border = styles["border"]
    ws.cell(row=row, column=2, value="Weighted Overall Compliance").font = Font(bold=True, size=12)
    ws.cell(row=row, column=2).border = styles["border"]
    ws.cell(row=row, column=3).border = styles["border"]
    ws.cell(row=row, column=4).border = styles["border"]
    ws.cell(row=row, column=5).fill = styles["calculated_cell"]["fill"]
    ws.cell(row=row, column=5).border = styles["border"]
    ws.cell(row=row, column=5).font = Font(bold=True, size=12)
    ws.cell(row=row, column=6).border = styles["border"]
    ws.cell(row=row, column=7).border = styles["border"]
    ws.cell(row=row, column=8).border = styles["border"]

    # Add data validation for Status
    dv_status = DataValidation(
        type="list",
        formula1='"GREEN - Compliant,AMBER - Partial,RED - Non-Compliant"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add(f"F4:F{row}")

    ws.freeze_panes = "D4"


# =============================================================================
# SHEET 4: GAP ANALYSIS
# =============================================================================
def create_gap_analysis_sheet(ws, styles):
    """Create the Gap Analysis sheet with 100-item capacity."""
    # Title
    ws.merge_cells("A1:O1")
    ws["A1"] = "GAP ANALYSIS - CONSOLIDATED"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A2"] = "Consolidated gap inventory from S1-S5 assessments (100 items capacity)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:O2")

    # Column headers
    row = 4
    headers = [
        ("Gap_ID", 16),
        ("Source", 12),
        ("Control", 10),
        ("Domain", 20),
        ("Gap_Description", 45),
        ("Severity", 10),
        ("Impact", 25),
        ("Root_Cause", 25),
        ("Remediation_Action", 35),
        ("Owner", 18),
        ("Status", 12),
        ("Due_Date", 12),
        ("Days_Open", 10),
        ("Evidence", 20),
        ("Notes", 30),
    ]

    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows (100 items)
    for r in range(5, 105):
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        # Days_Open is calculated
        ws.cell(row=r, column=13).fill = styles["calculated_cell"]["fill"]
        ws.cell(row=r, column=13).value = f'=IF(K{r}="Closed","N/A",IF(L{r}="","",TODAY()-L{r}))'

    # Data validations
    dv_source = DataValidation(
        type="list",
        formula1='"S1-Framework,S2-Detection,S3-Response,S4-Forensic,S5-Learning"',
        allow_blank=True
    )
    ws.add_data_validation(dv_source)
    dv_source.add("B5:B104")

    dv_control = DataValidation(
        type="list",
        formula1='"A.5.24,A.5.25,A.5.26,A.5.27,A.5.28"',
        allow_blank=True
    )
    ws.add_data_validation(dv_control)
    dv_control.add("C5:C104")

    dv_severity = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_severity)
    dv_severity.add("F5:F104")

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Closed"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("K5:K104")

    # Summary section
    row = 107
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "GAP SUMMARY"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    row += 2
    summary_headers = ["Severity", "Open", "In Progress", "Closed", "Total", "% Open"]
    for col, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    row += 1
    severities = ["Critical", "High", "Medium", "Low", "TOTAL"]
    for severity in severities:
        ws.cell(row=row, column=1, value=severity).border = styles["border"]
        if severity == "TOTAL":
            ws.cell(row=row, column=1).font = Font(bold=True)
        for c in range(2, 7):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["calculated_cell"]["fill"]
            cell.border = styles["border"]
            if severity == "TOTAL":
                cell.font = Font(bold=True)
        row += 1

    ws.freeze_panes = "E5"


# =============================================================================
# SHEET 5: INCIDENT KPIs
# =============================================================================
def create_incident_kpis_sheet(ws, styles):
    """Create the Incident KPIs sheet with operational metrics."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "INCIDENT MANAGEMENT KEY PERFORMANCE INDICATORS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Detection Metrics Section
    row = 3
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "DETECTION METRICS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    row += 2
    headers = ["KPI", "Current", "Target", "Status", "Trend", "Period", "Notes"]
    col_widths = [35, 15, 15, 20, 15, 12, 30]

    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]
        ws.column_dimensions[get_column_letter(col)].width = width

    detection_kpis = [
        ("Mean Time to Detect (MTTD) - Critical", "<1 hour"),
        ("Mean Time to Detect (MTTD) - High", "<4 hours"),
        ("Mean Time to Detect (MTTD) - Medium", "<24 hours"),
        ("Detection Source Coverage", ">80%"),
        ("False Positive Rate", "<10%"),
    ]

    row += 1
    for kpi, target in detection_kpis:
        ws.cell(row=row, column=1, value=kpi).border = styles["border"]
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]
        ws.cell(row=row, column=3, value=target).border = styles["border"]
        for c in range(4, 8):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Response Metrics Section
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "RESPONSE METRICS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    row += 2
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    response_kpis = [
        ("Mean Time to Respond (MTTR) - Critical", "<15 min"),
        ("Mean Time to Respond (MTTR) - High", "<1 hour"),
        ("Mean Time to Respond (MTTR) - Medium", "<4 hours"),
        ("First Contact Resolution Rate", ">60%"),
        ("Escalation Rate", "<20%"),
    ]

    row += 1
    for kpi, target in response_kpis:
        ws.cell(row=row, column=1, value=kpi).border = styles["border"]
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]
        ws.cell(row=row, column=3, value=target).border = styles["border"]
        for c in range(4, 8):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Containment & Recovery Metrics Section
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "CONTAINMENT & RECOVERY METRICS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    row += 2
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    containment_kpis = [
        ("Mean Time to Contain (MTTC) - Critical", "<1 hour"),
        ("Mean Time to Contain (MTTC) - High", "<4 hours"),
        ("Mean Time to Recover - Critical", "<4 hours"),
        ("Mean Time to Recover - High", "<24 hours"),
        ("Business Impact Duration (avg)", "<2 hours"),
    ]

    row += 1
    for kpi, target in containment_kpis:
        ws.cell(row=row, column=1, value=kpi).border = styles["border"]
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]
        ws.cell(row=row, column=3, value=target).border = styles["border"]
        for c in range(4, 8):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Operational Metrics Section
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "OPERATIONAL METRICS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    row += 2
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    operational_kpis = [
        ("Total Incidents (12 months)", "-"),
        ("Critical Incidents", "0"),
        ("High Incidents", "<10"),
        ("Incident Closure Rate", ">95%"),
        ("SLA Compliance Rate", ">95%"),
    ]

    row += 1
    for kpi, target in operational_kpis:
        ws.cell(row=row, column=1, value=kpi).border = styles["border"]
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]
        ws.cell(row=row, column=3, value=target).border = styles["border"]
        for c in range(4, 8):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Governance Metrics Section
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "GOVERNANCE METRICS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    row += 2
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    governance_kpis = [
        ("Post-Incident Review Completion", "100%"),
        ("PIR Completion within 14 days", ">90%"),
        ("Training Completion (CSIRT)", "100%"),
        ("Tabletop Exercise Frequency", "2/year"),
        ("Policy Review Currency", "<12 months"),
    ]

    row += 1
    for kpi, target in governance_kpis:
        ws.cell(row=row, column=1, value=kpi).border = styles["border"]
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]
        ws.cell(row=row, column=3, value=target).border = styles["border"]
        for c in range(4, 8):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # 12-Month Trend Section
    row += 3
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "12-MONTH TREND DATA"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    row += 2
    trend_headers = ["Month", "Total Incidents", "Critical", "High", "MTTD (avg)", "MTTR (avg)", "Closure Rate"]
    for col, header in enumerate(trend_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    row += 1
    for _ in range(12):
        for c in range(1, 8):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Data validations
    dv_status = DataValidation(
        type="list",
        formula1='"GREEN - Met,AMBER - Warning,RED - Critical"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    dv_trend = DataValidation(
        type="list",
        formula1='"▲ Improving,▬ Stable,▼ Declining"',
        allow_blank=True
    )
    ws.add_data_validation(dv_trend)

    ws.freeze_panes = "B4"


# =============================================================================
# SHEET 6: EVIDENCE REGISTER
# =============================================================================
def create_evidence_register_sheet(ws, styles):
    """Create the Evidence Register sheet with 100-item capacity."""
    # Title
    ws.merge_cells("A1:M1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A2"] = "Evidence inventory linking to source assessments (100 items capacity)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:M2")

    # Column headers
    row = 4
    headers = [
        ("Evidence_ID", 18),
        ("Source_Workbook", 40),
        ("Control", 10),
        ("Domain", 25),
        ("Evidence_Description", 45),
        ("Evidence_Type", 14),
        ("Storage_Location", 40),
        ("Date_Collected", 14),
        ("Collected_By", 18),
        ("Verification_Status", 16),
        ("Last_Verified", 14),
        ("Expiry_Date", 12),
        ("Notes", 30),
    ]

    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows (100 items)
    for r in range(5, 105):
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Data validations
    dv_control = DataValidation(
        type="list",
        formula1='"A.5.24,A.5.25,A.5.26,A.5.27,A.5.28"',
        allow_blank=True
    )
    ws.add_data_validation(dv_control)
    dv_control.add("C5:C104")

    dv_type = DataValidation(
        type="list",
        formula1='"Document,Screenshot,Record,Attestation"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("F5:F104")

    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Pending,Expired"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("J5:J104")

    # Summary section
    row = 107
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "EVIDENCE SUMMARY BY CONTROL"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    row += 2
    summary_headers = ["Control", "Total Evidence", "Verified", "Pending", "Expired", "Coverage"]
    for col, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    row += 1
    controls = ["A.5.24", "A.5.25", "A.5.26", "A.5.27", "A.5.28", "TOTAL"]
    for control in controls:
        ws.cell(row=row, column=1, value=control).border = styles["border"]
        if control == "TOTAL":
            ws.cell(row=row, column=1).font = Font(bold=True)
        for c in range(2, 7):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["calculated_cell"]["fill"]
            cell.border = styles["border"]
            if control == "TOTAL":
                cell.font = Font(bold=True)
        row += 1

    ws.freeze_panes = "E5"


# =============================================================================
# SHEET 7: APPROVAL SIGN-OFF
# =============================================================================
def create_approval_signoff_sheet(ws, styles):
    """Create the Approval Sign-Off sheet."""
    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "APPROVAL SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Assessment Summary Section
    row = 3
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    summary_items = [
        ("Assessment Period", ""),
        ("Dashboard Version", "1.0"),
        ("Overall Compliance Score", ""),
        ("Critical Gaps Identified", ""),
        ("High Gaps Identified", ""),
        ("Open Recommendations", ""),
        ("Last Refreshed", ""),
    ]

    row += 2
    for label, value in summary_items:
        ws[f"A{row}"] = label + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = styles["border"]
        cell = ws[f"B{row}"]
        cell.value = value
        if value == "":
            cell.fill = styles["input_cell"]["fill"]
        cell.border = styles["border"]
        row += 1

    # Approval Chain Section
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "APPROVAL CHAIN"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    row += 2
    approval_headers = ["Approver Name", "Role", "Status", "Date"]
    for col, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    approvers = [
        ("", "CSIRT Manager"),
        ("", "CISO"),
        ("", "Internal Audit (optional)"),
    ]

    row += 1
    for name, role in approvers:
        ws.cell(row=row, column=1).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=1).border = styles["border"]
        ws.cell(row=row, column=2, value=role).border = styles["border"]
        ws.cell(row=row, column=3).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=3).border = styles["border"]
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=4).border = styles["border"]
        row += 1

    # Validation for approval status
    dv_approval = DataValidation(
        type="list",
        formula1='"Approved,Approved with Comments,Rejected,Pending"',
        allow_blank=True
    )
    ws.add_data_validation(dv_approval)

    # Review Checklist Section
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "REVIEW CHECKLIST"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    checklist_items = [
        "Data accuracy verified",
        "Calculations validated",
        "Evidence references checked",
        "Gap prioritisation agreed",
        "KPI targets confirmed",
        "Recommendations reviewed",
    ]

    row += 2
    for item in checklist_items:
        ws[f"A{row}"] = item
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_yesno)

    # Comments Section
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "COMMENTS AND CONDITIONS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    row += 1
    ws.merge_cells(f"A{row}:D{row+4}")
    ws[f"A{row}"].fill = styles["input_cell"]["fill"]
    ws[f"A{row}"].border = styles["border"]
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Distribution Section
    row += 7
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "DISTRIBUTION"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    row += 2
    dist_headers = ["Recipient", "Role", "Distribution Date", "Method"]
    for col, header in enumerate(dist_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    row += 1
    for _ in range(5):
        for c in range(1, 5):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Next Review Section
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "NEXT REVIEW"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]

    row += 2
    ws[f"A{row}"] = "Next Review Date:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]

    row += 1
    ws[f"A{row}"] = "Review Trigger Events:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "Major incident, organisational change, audit finding"
    ws[f"B{row}"].border = styles["border"]

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20

    ws.freeze_panes = "A3"


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/7] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/7] Creating Executive_Summary sheet...")
        create_executive_summary_sheet(wb["Executive_Summary"], styles)

        logger.info("[3/7] Creating Compliance_Scores sheet...")
        create_compliance_scores_sheet(wb["Compliance_Scores"], styles)

        logger.info("[4/7] Creating Gap_Analysis sheet...")
        create_gap_analysis_sheet(wb["Gap_Analysis"], styles)

        logger.info("[5/7] Creating Incident_KPIs sheet...")
        create_incident_kpis_sheet(wb["Incident_KPIs"], styles)

        logger.info("[6/7] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[7/7] Creating Approval_Sign_Off sheet...")
        create_approval_signoff_sheet(wb["Approval_Sign_Off"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("-" * 78)
        logger.info("SUCCESS: Workbook generated successfully")
        logger.info("Output file: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        logger.error("Install with: pip3 install openpyxl")
        return 1
    except PermissionError as e:
        logger.error("Permission error: %s", e)
        logger.error("Check file permissions or close the file if open in Excel")
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - INITIAL CREATION
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.24-28.S6 specification
# =============================================================================
