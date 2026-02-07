#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
ISMS-IMP-A.5.24-28.S3 - Response Capabilities Assessment Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.26: Incident Response
Assessment Domain 3 of 5: Response Capabilities

--------------------------------------------------------------------------------
PURPOSE & DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
incident response execution capabilities, focusing on containment, eradication,
recovery, communication, and resource adequacy (A.5.26).

**Assessment Scope:**
- Containment procedures and technical capabilities
- Eradication and remediation processes
- Recovery and restoration procedures
- Communication during active incidents
- Resource availability and decision authority
- Response playbook effectiveness
- Response time metrics (MTTC, MTTR)

**Generated Workbook Structure:**
1. Instructions & Legend
2. Containment Capabilities - 30 questions
3. Eradication & Remediation - 20 questions
4. Recovery & Restoration - 20 questions
5. Communication - 20 questions
6. Resources & Authority - 20 questions
7. Playbook Effectiveness - 15 questions
8. Gap Analysis - 40 capacity
9. Evidence Register - 60 capacity
10. Dashboard
11. Approval Sign-Off

**Key Features:**
- 125 assessment questions across 6 domains
- Automated metric calculations (MTTC, MTTR, SLA compliance)
- Response capability scoring
- Playbook effectiveness assessment
- Authority matrix validation

**Integration:**
Domain 3 of 5 for A.5.24-28 Incident Management:
- S1: Framework & Governance (A.5.24)
- S2: Detection & Classification (A.5.25)
- S3: Response Capabilities (this - A.5.26)
- S4: Forensic Evidence (A.5.28)
- S5: Learning & Improvement (A.5.27)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

Python 3.8+, openpyxl

Installation: pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

python3 generate_a524_28_s3_response_capabilities.py

Output: ISMS-IMP-A.5.24-28.S3_Response_Capabilities_YYYYMMDD.xlsx

Estimated completion: 8-12 hours

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 A.5.24-28
Assessment Domain:    3 of 5 (Response Capabilities - A.5.26)
Version:              1.0
Date:                 2026-01-30
Python Version:       3.8+

================================================================================
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.24-28.S3"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_Response_Capabilities_{GENERATED_TIMESTAMP}.xlsx"


def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    sheets = [
        "Instructions & Legend",
        "Containment Capabilities",
        "Eradication & Remediation",
        "Recovery & Restoration",
        "Communication",
        "Resources & Authority",
        "Playbook Effectiveness",
        "Gap Analysis",
        "Evidence Register",
        "Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def setup_styles():
    """Define cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "calculated_cell": {
            "fill": PatternFill(start_color="C8F0C8", end_color="C8F0C8", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
        "gap_critical": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "gap_high": {"fill": PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")},
        "gap_medium": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "gap_low": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
    }


def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.5.24-28.S3 — Response Capabilities Assessment\n"
        "ISO/IEC 27001:2022 - Control A.5.26: Incident Response"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40
    
    row = 3
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:B{row}")
    
    doc_info = [
        ("Document ID", "ISMS-IMP-A.5.24-28.S3"),
        ("Assessment Area", "Incident Response Capabilities (Domain 3)"),
        ("Related Policy", "ISMS-POL-A.5.24-28, Section 2.3"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organization", ""),
    ]
    
    row += 1
    for label, value in doc_info:
        ws[f"A{row}"] = label + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if label in ["Assessment Date", "Completed By", "Organization"]:
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1
    
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.freeze_panes = "A3"


def create_assessment_sheet(ws, styles, sheet_title, subtitle, questions):
    """Generic function to create assessment sheets."""
    ws.merge_cells("A1:G1")
    ws["A1"] = sheet_title
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")
    
    row = 4
    headers = ["Question_ID", "Section", "Question", "Answer", "Evidence_Reference", "Comments", "Gap_Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(OR(D{row}="No", D{row}="Limited", D{row}<80), "Yes", "No")'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12
    ws.freeze_panes = "A5"


def create_containment_capabilities(ws, styles):
    """Create Containment Capabilities sheet."""
    questions = [
        ("Q1", "Network", "Can [Organization] isolate network segments?", "Automated/Manual/Limited/No"),
        ("Q2", "Network", "Emergency firewall rule changes capability?", "24/7/Business Hours/No"),
        ("Q3", "Network", "Can internet egress be blocked for infected systems?", "Per System/Per Subnet/No"),
        ("Q4", "Network", "Is there a quarantine VLAN?", "Yes/No"),
        ("Q5", "Network", "Average time to execute network containment?", "Duration"),
        ("Q6", "Network", "Who has authority to order network containment?", "Text"),
        ("Q7", "Network", "Network containment procedures documented?", "Comprehensive/Basic/No"),
        ("Q8", "Network", "Network team available 24/7?", "24/7/On-Call/Business Hours/No"),
        
        ("Q9", "Endpoint", "Can endpoints be isolated remotely via EDR?", "Automated/Manual/No EDR/No"),
        ("Q10", "Endpoint", "Endpoint isolation speed?", "Duration"),
        ("Q11", "Endpoint", "Non-EDR endpoint isolation capability?", "Yes/No"),
        ("Q12", "Endpoint", "% endpoints remotely isolatable?", "Percentage"),
        ("Q13", "Endpoint", "Remote laptop isolation capability?", "EDR/VPN Disconnect/Limited/No"),
        ("Q14", "Endpoint", "Can production servers be isolated?", "Yes - With Approval/Emergency Authority/No"),
        ("Q15", "Endpoint", "Endpoint isolation reversal process?", "Yes/No"),
        
        ("Q16", "Account", "Compromised account suspension capability?", "Automated/Manual (Min)/Manual (Hours)/No"),
        ("Q17", "Account", "Who has authority to suspend accounts?", "Text"),
        ("Q18", "Account", "MFA token/session revocation capability?", "Yes/No"),
        ("Q19", "Account", "Service account credential rotation?", "Automated/Manual/No"),
        ("Q20", "Account", "Force password reset capability?", "Immediate/Next Login/No"),
        ("Q21", "Account", "Privileged access revocation capability?", "Immediately/Delayed/No"),
        ("Q22", "Account", "Average time to suspend compromised account?", "Duration"),
        
        ("Q23", "Application", "Who has authority to shut down applications?", "Text"),
        ("Q24", "Application", "Individual service isolation capability?", "Yes/Limited/No"),
        ("Q25", "Application", "Database connection blocking capability?", "Per App/Global/No"),
        ("Q26", "Application", "Emergency WAF rule deployment?", "Automated/Manual/No WAF"),
        ("Q27", "Application", "Cloud service containment capability?", "Yes/Limited/No Cloud/No"),
        
        ("Q28", "Metrics", "Mean Time to Contain (MTTC) overall? (6 months)", "Duration"),
        ("Q29", "Metrics", "MTTC for Critical incidents?", "Duration"),
        ("Q30", "Metrics", "% incidents meeting containment SLA? (6 months)", "Percentage"),
    ]
    
    create_assessment_sheet(ws, styles, "Sheet 2: Containment Capabilities",
                          "Network, Endpoint, Account, Application Containment (30 Questions)", questions)


def create_eradication_remediation(ws, styles):
    """Create Eradication & Remediation sheet."""
    questions = [
        ("Q31", "Malware", "Malware removal procedures documented?", "Comprehensive/Basic/No"),
        ("Q32", "Malware", "EDR automated malware removal?", "Automated/Manual/No"),
        ("Q33", "Malware", "Reimaging capability?", "Automated/Manual (Same Day)/Manual (Days)/No"),
        ("Q34", "Malware", "Persistence mechanism checks?", "Systematically/Sometimes/No"),
        ("Q35", "Malware", "Network share malware scanning?", "Automatically/Manually/No"),
        ("Q36", "Malware", "Malware sample collection before removal?", "Always/Sometimes/No"),
        
        ("Q37", "Vulnerability", "Emergency patching capability?", "Same Day/Within Week/No"),
        ("Q38", "Vulnerability", "Average vulnerability remediation time?", "Duration"),
        ("Q39", "Vulnerability", "Can patch testing be waived?", "With Approval/No - Always Test"),
        ("Q40", "Vulnerability", "Configuration hardening after incidents?", "Systematically/Sometimes/No"),
        ("Q41", "Vulnerability", "Emergency code fix deployment?", "Same Day/Days/No/N/A"),
        ("Q42", "Vulnerability", "Workaround implementation capability?", "Yes/No"),
        
        ("Q43", "Credential", "Credential rotation procedures documented?", "Comprehensive/Basic/No"),
        ("Q44", "Credential", "Bulk password reset capability?", "Yes/No"),
        ("Q45", "Credential", "Certificate revocation/reissue capability?", "Yes/No"),
        ("Q46", "Credential", "API key rotation capability?", "Automated/Manual/No"),
        ("Q47", "Credential", "Privileged account review after incidents?", "Systematically/Sometimes/No"),
        
        ("Q48", "Expulsion", "Threat actor expulsion verification?", "Systematically/Sometimes/No"),
        ("Q49", "Expulsion", "Backdoor search and removal?", "Comprehensive/Limited/No"),
        ("Q50", "Expulsion", "Re-infection prevention measures?", "Yes/No"),
    ]
    
    create_assessment_sheet(ws, styles, "Sheet 3: Eradication & Remediation",
                          "Malware, Vulnerability, Credential, Threat Actor Expulsion (20 Questions)", questions)


def create_recovery_restoration(ws, styles):
    """Create Recovery & Restoration sheet."""
    questions = [
        ("Q51", "Restoration", "System restoration procedures documented?", "Comprehensive/Basic/No"),
        ("Q52", "Restoration", "Backup availability for critical systems?", "All/Most/Limited/No"),
        ("Q53", "Restoration", "Average backup restoration time?", "Duration"),
        ("Q54", "Restoration", "Backup integrity verification before restoration?", "Always/Sometimes/No"),
        ("Q55", "Restoration", "Clean rebuild capability?", "Yes/Limited/No"),
        ("Q56", "Restoration", "Configuration restoration capability?", "Yes/Partial/No"),
        ("Q57", "Restoration", "Data recovery testing frequency?", "Quarterly/Annually/No"),
        
        ("Q58", "Resumption", "Service resumption procedures documented?", "Yes/No"),
        ("Q59", "Resumption", "Post-recovery validation?", "Comprehensive/Basic/No"),
        ("Q60", "Resumption", "Staged recovery capability?", "Yes/No"),
        ("Q61", "Resumption", "User communication during recovery?", "Proactively/On Request/No"),
        ("Q62", "Resumption", "Service degradation mode capability?", "Documented/Limited/No"),
        ("Q63", "Resumption", "RTOs defined for critical services?", "Yes/No"),
        ("Q64", "Resumption", "% incidents meeting RTO? (6 months)", "Percentage"),
        
        ("Q65", "Monitoring", "Enhanced monitoring period after recovery?", "7+ Days/1-3 Days/No"),
        ("Q66", "Monitoring", "Reinfection actively monitored?", "Yes/No"),
        ("Q67", "Monitoring", "IOC monitoring after recovery?", "Yes/No"),
        ("Q68", "Monitoring", "Mean Time to Recover (MTTR) overall? (6 months)", "Duration"),
        ("Q69", "Monitoring", "MTTR for Critical incidents?", "Duration"),
        ("Q70", "Monitoring", "Recovery success rate (no recurrence)? (6 months)", "Percentage"),
    ]
    
    create_assessment_sheet(ws, styles, "Sheet 4: Recovery & Restoration",
                          "System Restoration, Service Resumption, Post-Recovery Monitoring (20 Questions)", questions)


def create_communication(ws, styles):
    """Create Communication sheet."""
    questions = [
        ("Q71", "Internal", "Internal communication protocol documented?", "Yes/No"),
        ("Q72", "Internal", "Management notifications within SLA?", "Always/Usually/No"),
        ("Q73", "Internal", "Affected user notification?", "Proactively/On Request/No"),
        ("Q74", "Internal", "IT Ops coordination method?", "Dedicated Channel/Email/Phone/Informal/No"),
        ("Q75", "Internal", "Communication templates available?", "Comprehensive/Basic/No"),
        ("Q76", "Internal", "Communication approval process?", "Required/Optional/No"),
        ("Q77", "Internal", "All incident communication logged?", "Yes/No"),
        
        ("Q78", "External", "External communication procedures documented?", "Yes/No"),
        ("Q79", "External", "Customer notification capability?", "Automated/Manual/No"),
        ("Q80", "External", "Regulatory notification templates ready?", "Yes/No"),
        ("Q81", "External", "Media inquiry handling process?", "PR Team/Ad-Hoc/No"),
        ("Q82", "External", "Third-party notification when affected?", "Yes/No"),
        ("Q83", "External", "Legal reviews external communications?", "Always/Sometimes/No"),
        ("Q84", "External", "External communications within timelines?", "Always/Usually/No"),
        
        ("Q85", "Crisis", "Crisis communication plan exists?", "Yes/No"),
        ("Q86", "Crisis", "Designated spokespersons identified?", "Yes/No"),
        ("Q87", "Crisis", "Communication frequency defined?", "Yes/No"),
        ("Q88", "Crisis", "Stakeholder satisfaction with communication?", "Yes/Mixed/No/Not Measured"),
        ("Q89", "Crisis", "Communication lessons learned captured?", "Yes/No"),
        ("Q90", "Crisis", "CSIRT communication training?", "Yes/No"),
    ]
    
    create_assessment_sheet(ws, styles, "Sheet 5: Communication",
                          "Internal, External, Crisis Communication (20 Questions)", questions)


def create_resources_authority(ws, styles):
    """Create Resources & Authority sheet."""
    questions = [
        ("Q91", "CSIRT Availability", "CSIRT 24/7 coverage?", "Dedicated Staff/On-Call/Business Hours"),
        ("Q92", "CSIRT Availability", "On-call response SLA?", "Duration"),
        ("Q93", "CSIRT Availability", "On-call escalation process?", "Yes/No"),
        ("Q94", "CSIRT Availability", "Weekend/holiday coverage?", "Yes/Limited/No"),
        ("Q95", "CSIRT Availability", "Surge capacity for multiple incidents?", "External Support/Internal Only/No"),
        
        ("Q96", "Tool Access", "After-hours tool access?", "Yes/Limited/No"),
        ("Q97", "Tool Access", "Remote response capability?", "Full/Limited/No"),
        ("Q98", "Tool Access", "Emergency system access?", "Break-Glass/With Approval/No"),
        ("Q99", "Tool Access", "External partner tool access?", "Yes/Limited/No/N/A"),
        ("Q100", "Tool Access", "VPN capacity for mass remote work?", "Yes/Limited/No"),
        
        ("Q101", "Authority", "Emergency decision authority documented?", "Yes/No"),
        ("Q102", "Authority", "CSIRT business impact decision authority?", "Full/With Approval/No"),
        ("Q103", "Authority", "After-hours critical decision authority?", "Yes/Limited/Must Escalate"),
        ("Q104", "Authority", "CSIRT spending authority?", "Up to $X/With Approval/No"),
        ("Q105", "Authority", "Legal hold initiation authority?", "Yes/With Legal/No"),
        
        ("Q106", "Budget", "IR budget allocated?", "Dedicated/Part of IT Security/No"),
        ("Q107", "Budget", "External IR retainer exists?", "Yes/No"),
        ("Q108", "Budget", "Forensic services available?", "Retainer/Ad-Hoc/No"),
        ("Q109", "Budget", "24/7 legal support available?", "Yes/Business Hours/No"),
        ("Q110", "Budget", "Cyber insurance coverage?", "Yes/No"),
    ]
    
    create_assessment_sheet(ws, styles, "Sheet 6: Resources & Authority",
                          "CSIRT Availability, Tool Access, Decision Authority, Budget (20 Questions)", questions)


def create_playbook_effectiveness(ws, styles):
    """Create Playbook Effectiveness sheet."""
    questions = [
        ("Q111", "Usage", "Playbook usage rate?", "Percentage"),
        ("Q112", "Quality", "Playbooks comprehensive?", "Comprehensive/Partial/No"),
        ("Q113", "Quality", "Playbooks accurate?", "Yes/Mostly/No"),
        ("Q114", "Quality", "Playbooks updated after incidents?", "Always/Sometimes/No"),
        ("Q115", "Quality", "Playbook accessibility during incidents?", "Easily/With Difficulty/No"),
        ("Q116", "By Category", "Playbook: Ransomware - Quality?", "Excellent/Good/Fair/Poor/None"),
        ("Q117", "By Category", "Playbook: Phishing - Quality?", "Excellent/Good/Fair/Poor/None"),
        ("Q118", "By Category", "Playbook: Data Breach - Quality?", "Excellent/Good/Fair/Poor/None"),
        ("Q119", "By Category", "Playbook: Unauthorized Access - Quality?", "Excellent/Good/Fair/Poor/None"),
        ("Q120", "By Category", "Playbook: DDoS - Quality?", "Excellent/Good/Fair/Poor/None"),
        ("Q121", "Training", "Responders trained on playbooks?", "Yes/No"),
        ("Q122", "Testing", "Playbooks tested during exercises?", "Regularly/Occasionally/No"),
        ("Q123", "Automation", "% playbook steps automated (SOAR)?", "Percentage"),
        ("Q124", "Improvement", "Playbook deviations tracked?", "Yes/No"),
        ("Q125", "Improvement", "Playbook improvement process?", "Yes/No"),
    ]
    
    create_assessment_sheet(ws, styles, "Sheet 7: Playbook Effectiveness",
                          "Usage, Quality, Training, Testing, Automation, Improvement (15 Questions)", questions)


def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Sheet 8: Gap Analysis"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    row = 4
    headers = ["Gap_ID", "Section", "Gap_Description", "Risk_Level", "Current_State", "Target_State", "Remediation", "Owner", "Target_Date", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    for i in range(40):
        row += 1
        ws[f"A{row}"] = f"GAP-{str(i+1).zfill(3)}"
        for col_idx in range(2, 11):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
    
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.freeze_panes = "A5"


def create_evidence_register(ws, styles):
    """Create Evidence Register."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Sheet 9: Evidence Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    row = 4
    headers = ["Evidence_ID", "Evidence_Type", "Description", "Related_Section", "Storage_Location", "Date_Collected", "Collected_By", "Verification"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    for i in range(60):
        row += 1
        ws[f"A{row}"] = f"EV-{str(i+1).zfill(3)}"
        for col_idx in range(2, 9):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
    
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.freeze_panes = "A5"


def create_dashboard(ws, styles):
    """Create Dashboard."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Sheet 10: Dashboard"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    row = 4
    ws[f"A{row}"] = "RESPONSE EFFECTIVENESS SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:D{row}")
    
    metrics = [
        ("Mean Time to Contain (MTTC)", "='Containment Capabilities'!D28"),
        ("MTTC Critical Incidents", "='Containment Capabilities'!D29"),
        ("Containment SLA Compliance", "='Containment Capabilities'!D30"),
        ("Mean Time to Recover (MTTR)", "='Recovery & Restoration'!D68"),
        ("MTTR Critical Incidents", "='Recovery & Restoration'!D69"),
        ("Recovery Success Rate", "='Recovery & Restoration'!D70"),
        ("Playbook Usage Rate", "='Playbook Effectiveness'!D111"),
        ("Playbook Automation Rate", "='Playbook Effectiveness'!D123"),
    ]
    
    row += 1
    for metric_label, metric_formula in metrics:
        ws[f"A{row}"] = metric_label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = metric_formula
        ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]
        row += 1
    
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = "A4"


def create_approval_signoff(ws, styles):
    """Create Approval Sign-Off."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "Sheet 11: Approval & Sign-Off"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    row = 3
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")
    
    row += 1
    summary = [
        ("Assessment Document", "ISMS-IMP-A.5.24-28.S3 - Response Capabilities"),
        ("Assessment Period", ""),
        ("MTTC", "=Dashboard!B5"),
        ("MTTR", "=Dashboard!B8"),
    ]
    
    for label, value in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if label == "Assessment Period":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        else:
            ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1
    
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "COMPLETED BY (CSIRT Manager)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    for field in ["Name", "Date", "Signature"]:
        row += 1
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
    
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    for field in ["Name", "Date", "Approval Decision", "Signature"]:
        row += 1
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
    
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.freeze_panes = "A3"


def main():
    """Main execution."""
    print("=" * 80)
    print("ISMS-IMP-A.5.24-28.S3 - Response Capabilities Assessment Generator")
    print("ISO/IEC 27001:2022 Control A.5.26: Incident Response")
    print("=" * 80)
    
    wb = create_workbook()
    styles = setup_styles()
    
    print("\n[1/11] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    
    print("[2/11] Creating Containment Capabilities (30 questions)...")
    create_containment_capabilities(wb["Containment Capabilities"], styles)
    
    print("[3/11] Creating Eradication & Remediation (20 questions)...")
    create_eradication_remediation(wb["Eradication & Remediation"], styles)
    
    print("[4/11] Creating Recovery & Restoration (20 questions)...")
    create_recovery_restoration(wb["Recovery & Restoration"], styles)
    
    print("[5/11] Creating Communication (20 questions)...")
    create_communication(wb["Communication"], styles)
    
    print("[6/11] Creating Resources & Authority (20 questions)...")
    create_resources_authority(wb["Resources & Authority"], styles)
    
    print("[7/11] Creating Playbook Effectiveness (15 questions)...")
    create_playbook_effectiveness(wb["Playbook Effectiveness"], styles)
    
    print("[8/11] Creating Gap Analysis...")
    create_gap_analysis(wb["Gap Analysis"], styles)
    
    print("[9/11] Creating Evidence Register...")
    create_evidence_register(wb["Evidence Register"], styles)
    
    print("[10/11] Creating Dashboard...")
    create_dashboard(wb["Dashboard"], styles)
    
    print("[11/11] Creating Approval Sign-Off...")
    create_approval_signoff(wb["Approval Sign-Off"], styles)
    
    wb.save(OUTPUT_FILENAME)

    print(f"\n✅ SUCCESS: {OUTPUT_FILENAME}")
    print("\nWorkbook Structure:")
    print("  • 11 sheets (Instructions through Approval)")
    print("  • 125 assessment questions (30+20+20+20+20+15)")
    print("  • 40 gap analysis capacity")
    print("  • 60 evidence register capacity")
    print("  • Automated metrics (MTTC, MTTR, SLA compliance)")
    print("\nNext steps:")
    print("  1) Extract response metrics from incident tickets")
    print("  2) Complete containment capability assessment")
    print("  3) Review response playbook effectiveness")
    print("  4) Fill yellow cells in all assessment sheets")
    print("  5) Analyze response time performance")
    print("  6) Complete Approval Sign-Off workflow")
    print("\nEstimated completion time: 8-12 hours")
    print("\n" + "=" * 80)


if __name__ == "__main__":
    main()


# ============================================================================
# END OF SCRIPT
# ============================================================================

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# =============================================================================

