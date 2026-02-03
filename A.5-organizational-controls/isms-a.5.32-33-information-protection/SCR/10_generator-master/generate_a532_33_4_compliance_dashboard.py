#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.32-33.4 - Information Protection Compliance Dashboard Generator
================================================================================

ISO/IEC 27001:2022 Controls A.5.32 & A.5.33
Assessment Domain 4 of 4: Compliance Dashboard

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive compliance dashboard for monitoring
the organisation's information protection program effectiveness.

**Purpose:**
Provide executive-level visibility into IP protection, records protection,
and retention compliance metrics.

**Generated Workbook Structure:**
1. Instructions - Dashboard usage guidance
2. Executive_Summary - Key metrics and compliance status
3. Compliance_Metrics - Detailed KPIs
4. Control_Assessment - A.5.32 and A.5.33 evaluation
5. Maturity_Assessment - Program maturity scoring
6. Risk_Register - Information protection risks
7. Remediation_Tracker - Gap closure tracking
8. Trend_Analysis - Period over period comparison
9. Evidence_Register - Audit evidence tracking
10. Approval_SignOff - Dashboard approval

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.32-33.4"
WORKBOOK_NAME = "Information Protection Compliance Dashboard"
CONTROL_ID = "A.5.32-33"
CONTROL_NAME = "Intellectual Property Rights & Protection of Records"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "metric_good": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "metric_warning": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "metric_bad": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "border": border_thin,
    }


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    for attr, value in style_dict.items():
        if attr != "border":
            setattr(cell, attr, value)


def set_column_widths(ws, widths):
    """Set column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    apply_style(ws["A2"], styles["subheader"])

    instructions = [
        "",
        "PURPOSE",
        "This dashboard consolidates metrics from all A.5.32-33 assessments to provide",
        "executive visibility into information protection compliance status.",
        "",
        "DATA SOURCES",
        "- A.5.32-33.1: IP Rights Inventory (IP asset count, protection status)",
        "- A.5.32-33.2: Records Protection (control effectiveness, integrity)",
        "- A.5.32-33.3: Retention & Disposal (compliance rate, disposal completion)",
        "",
        "UPDATE FREQUENCY",
        "- Executive Summary: Quarterly",
        "- Compliance Metrics: Quarterly",
        "- Control Assessment: Annual",
        "- Maturity Assessment: Annual",
        "- Risk Register: Quarterly",
        "- Remediation Tracker: Monthly",
        "",
        "KEY STAKEHOLDERS",
        "- Executive Management: Overall compliance status",
        "- CISO: Technical control effectiveness",
        "- Legal Counsel: IP protection and regulatory compliance",
        "- Records Manager: Retention compliance",
        "- Internal Audit: Control assessment verification",
        "",
        "HOW TO USE THIS DASHBOARD",
        "1. Gather metrics from source assessments (A.5.32-33.1, .2, .3)",
        "2. Update Executive_Summary with overall status",
        "3. Calculate and enter Compliance_Metrics",
        "4. Complete Control_Assessment for A.5.32 and A.5.33",
        "5. Update Maturity_Assessment scores",
        "6. Refresh Risk_Register with current risks",
        "7. Update Remediation_Tracker progress",
        "8. Analyse trends in Trend_Analysis",
        "9. Obtain stakeholder approval",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, line in enumerate(instructions, start=4):
        ws[f"A{i}"] = line
        if line in ["PURPOSE", "DATA SOURCES", "UPDATE FREQUENCY",
                    "KEY STAKEHOLDERS", "HOW TO USE THIS DASHBOARD"]:
            ws[f"A{i}"].font = Font(bold=True, size=11)

    set_column_widths(ws, {"A": 100})
    logger.info("Created Instructions sheet")


def create_executive_summary_sheet(ws, styles):
    """Create the Executive Summary sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{DOCUMENT_ID} - Executive Summary"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    apply_style(ws["A2"], styles["subheader"])

    # Assessment period
    ws["A4"] = "ASSESSMENT INFORMATION"
    ws["A4"].font = Font(bold=True, size=12)

    ws["A5"] = "Assessment Period:"
    ws["B5"] = ""
    ws["A6"] = "Review Date:"
    ws["B6"] = GENERATED_DATE
    ws["A7"] = "Next Review Date:"
    ws["B7"] = ""

    for row in range(5, 8):
        ws[f"A{row}"].font = Font(bold=True)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]

    # Overall Status
    ws["A10"] = "OVERALL COMPLIANCE STATUS"
    ws["A10"].font = Font(bold=True, size=12)

    status_items = [
        ("Overall Status", ""),
        ("IP Protection Status", ""),
        ("Records Protection Status", ""),
        ("Retention Compliance Status", ""),
    ]

    for i, (item, status) in enumerate(status_items, start=11):
        ws.cell(row=i, column=1, value=item).font = Font(bold=True)
        ws.cell(row=i, column=1).border = styles["border"]
        cell = ws.cell(row=i, column=2, value=status)
        cell.border = styles["border"]
        cell.fill = styles["input_cell"]["fill"]

    # Key Metrics Summary
    ws["A17"] = "KEY METRICS SUMMARY"
    ws["A17"].font = Font(bold=True, size=12)

    metrics_headers = ["Metric", "Value", "Target", "Status"]
    for col, header in enumerate(metrics_headers, start=1):
        cell = ws.cell(row=18, column=col, value=header)
        apply_style(cell, styles["column_header"])

    metrics = [
        ("IP Assets with Owners", "", "100%", ""),
        ("Third-Party IP Compliance", "", "100%", ""),
        ("Software License Compliance", "", "100%", ""),
        ("Records Protection Effectiveness", "", "90%", ""),
        ("Retention Schedule Compliance", "", "95%", ""),
        ("Disposal Completion Rate", "", "95%", ""),
    ]

    for i, (metric, value, target, status) in enumerate(metrics, start=19):
        ws.cell(row=i, column=1, value=metric).border = styles["border"]
        for col in [2, 3, 4]:
            cell = ws.cell(row=i, column=col, value=[value, target, status][col-2])
            cell.border = styles["border"]
            if col in [2, 4]:
                cell.fill = styles["input_cell"]["fill"]

    # Key Achievements and Concerns
    ws["A27"] = "KEY ACHIEVEMENTS"
    ws["A27"].font = Font(bold=True, size=12)
    ws.merge_cells("A28:D30")
    ws["A28"].fill = styles["input_cell"]["fill"]
    ws["A28"].border = styles["border"]
    ws["A28"].alignment = Alignment(vertical="top", wrap_text=True)

    ws["A32"] = "KEY CONCERNS"
    ws["A32"].font = Font(bold=True, size=12)
    ws.merge_cells("A33:D35")
    ws["A33"].fill = styles["input_cell"]["fill"]
    ws["A33"].border = styles["border"]
    ws["A33"].alignment = Alignment(vertical="top", wrap_text=True)

    # Data validations
    dv_overall = DataValidation(
        type="list",
        formula1='"Compliant,Partial,Non-Compliant"',
        allow_blank=True
    )
    ws.add_data_validation(dv_overall)
    dv_overall.add("B11")

    dv_status = DataValidation(
        type="list",
        formula1='"Effective,Partial,Ineffective"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("B12:B13")

    dv_retention = DataValidation(
        type="list",
        formula1='"Compliant,At Risk,Non-Compliant"',
        allow_blank=True
    )
    ws.add_data_validation(dv_retention)
    dv_retention.add("B14")

    dv_metric_status = DataValidation(
        type="list",
        formula1='"On Target,At Risk,Below Target"',
        allow_blank=True
    )
    ws.add_data_validation(dv_metric_status)
    dv_metric_status.add("D19:D24")

    set_column_widths(ws, {"A": 35, "B": 20, "C": 15, "D": 18})
    logger.info("Created Executive_Summary sheet")


def create_compliance_metrics_sheet(ws, styles):
    """Create the Compliance Metrics sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "Compliance Metrics"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = "Key Performance Indicators for information protection"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Metric ID", "Category", "Metric Name", "Description",
        "Target", "Current Value", "Previous Value", "Trend",
        "Status", "Owner", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Sample metrics
    metrics = [
        ["KPI-IP-001", "IP Protection", "IP Assets with Owners",
         "% of IP assets with designated owner", "100%", "", "", "", "", "Legal Counsel", ""],
        ["KPI-IP-002", "IP Protection", "IP Protection Effectiveness",
         "% of IP assets with effective controls", "95%", "", "", "", "", "CISO", ""],
        ["KPI-IP-003", "IP Protection", "Third-Party Compliance",
         "% of third-party IP properly licensed", "100%", "", "", "", "", "IT Ops", ""],
        ["KPI-IP-004", "IP Protection", "Software License Compliance",
         "% of software within license entitlements", "100%", "", "", "", "", "IT Ops", ""],
        ["KPI-REC-001", "Records Protection", "Records Classification Coverage",
         "% of record categories with retention defined", "100%", "", "", "", "", "Records Mgr", ""],
        ["KPI-REC-002", "Records Protection", "Protection Control Effectiveness",
         "% of controls rated effective", "90%", "", "", "", "", "CISO", ""],
        ["KPI-REC-003", "Records Protection", "Integrity Verification Pass Rate",
         "% of integrity tests passing", "100%", "", "", "", "", "IT Ops", ""],
        ["KPI-RET-001", "Retention/Disposal", "On-Schedule Disposal Rate",
         "% of records disposed within grace period", "95%", "", "", "", "", "Records Mgr", ""],
        ["KPI-RET-002", "Retention/Disposal", "Overdue Disposals",
         "Count of records past retention + grace", "<5%", "", "", "", "", "Records Mgr", ""],
        ["KPI-RET-003", "Retention/Disposal", "Destruction Certificate Rate",
         "% of disposals with certificates", "100%", "", "", "", "", "Records Mgr", ""],
    ]

    for row_idx, metric in enumerate(metrics, start=5):
        for col_idx, value in enumerate(metric, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx in [6, 7, 8, 9]:
                cell.fill = styles["input_cell"]["fill"]

    # Add empty rows
    for row in range(15, 30):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_category = DataValidation(
        type="list",
        formula1='"IP Protection,Records Protection,Retention/Disposal"',
        allow_blank=True
    )
    ws.add_data_validation(dv_category)
    dv_category.add("B5:B35")

    dv_trend = DataValidation(
        type="list",
        formula1='"Improving,Stable,Declining,New"',
        allow_blank=True
    )
    ws.add_data_validation(dv_trend)
    dv_trend.add("H5:H35")

    dv_status = DataValidation(
        type="list",
        formula1='"On Target,At Risk,Below Target"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("I5:I35")

    set_column_widths(ws, {
        "A": 12, "B": 18, "C": 28, "D": 40, "E": 10,
        "F": 12, "G": 15, "H": 12, "I": 15, "J": 15, "K": 25
    })
    logger.info("Created Compliance_Metrics sheet")


def create_control_assessment_sheet(ws, styles):
    """Create the Control Assessment sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "Control A.5.32 & A.5.33 Assessment"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "Detailed evaluation of ISO 27001:2022 control requirements"
    apply_style(ws["A2"], styles["subheader"])

    # A.5.32 Assessment
    ws["A4"] = "CONTROL A.5.32: INTELLECTUAL PROPERTY RIGHTS"
    ws["A4"].font = Font(bold=True, size=12)

    headers = ["Requirement", "Implementation", "Evidence", "Gap", "Score", "Status"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["column_header"])

    a532_reqs = [
        ["IP identification procedures implemented", "", "", "", "", ""],
        ["IP classification scheme defined", "", "", "", "", ""],
        ["Protection controls per IP category", "", "", "", "", ""],
        ["Third-party IP compliance verified", "", "", "", "", ""],
        ["Software licensing compliance", "", "", "", "", ""],
        ["IP ownership clearly assigned", "", "", "", "", ""],
        ["Periodic IP review process", "", "", "", "", ""],
    ]

    for row_idx, req in enumerate(a532_reqs, start=6):
        for col_idx, value in enumerate(req, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx > 1:
                cell.fill = styles["input_cell"]["fill"]

    # A.5.33 Assessment
    ws["A15"] = "CONTROL A.5.33: PROTECTION OF RECORDS"
    ws["A15"].font = Font(bold=True, size=12)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=16, column=col, value=header)
        apply_style(cell, styles["column_header"])

    a533_reqs = [
        ["Records inventory maintained", "", "", "", "", ""],
        ["Records classified by retention requirements", "", "", "", "", ""],
        ["Protection from loss (backups)", "", "", "", "", ""],
        ["Protection from destruction (controls)", "", "", "", "", ""],
        ["Protection from falsification (integrity)", "", "", "", "", ""],
        ["Protection from unauthorised access", "", "", "", "", ""],
        ["Protection from unauthorised release", "", "", "", "", ""],
        ["Retention schedule defined and followed", "", "", "", "", ""],
        ["Secure disposal procedures implemented", "", "", "", "", ""],
    ]

    for row_idx, req in enumerate(a533_reqs, start=17):
        for col_idx, value in enumerate(req, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx > 1:
                cell.fill = styles["input_cell"]["fill"]

    dv_score = DataValidation(
        type="list",
        formula1='"5 - Optimised,4 - Managed,3 - Defined,2 - Developing,1 - Initial,0 - Non-existent"',
        allow_blank=True
    )
    ws.add_data_validation(dv_score)
    dv_score.add("E6:E12")
    dv_score.add("E17:E25")

    dv_status = DataValidation(
        type="list",
        formula1='"Compliant,Partial,Non-Compliant"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("F6:F12")
    dv_status.add("F17:F25")

    set_column_widths(ws, {
        "A": 40, "B": 30, "C": 25, "D": 25, "E": 18, "F": 15
    })
    logger.info("Created Control_Assessment sheet")


def create_maturity_assessment_sheet(ws, styles):
    """Create the Maturity Assessment sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "Information Protection Maturity Assessment"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = "CMMI-based maturity model for information protection program"
    apply_style(ws["A2"], styles["subheader"])

    # Maturity scale reference
    ws["A4"] = "MATURITY SCALE REFERENCE"
    ws["A4"].font = Font(bold=True, size=11)

    scale = [
        ("5 - Optimised", "Continuous improvement, automated, metrics-driven"),
        ("4 - Managed", "Measured and controlled, consistent performance"),
        ("3 - Defined", "Documented, standardised across organisation"),
        ("2 - Developing", "Reactive, informal, department-specific"),
        ("1 - Initial", "Ad hoc, unpredictable, poorly controlled"),
        ("0 - Non-existent", "No awareness, no processes"),
    ]

    for i, (level, desc) in enumerate(scale, start=5):
        ws.cell(row=i, column=1, value=level).font = Font(bold=True)
        ws.cell(row=i, column=1).border = styles["border"]
        cell = ws.cell(row=i, column=2, value=desc)
        cell.border = styles["border"]
        ws.merge_cells(f"B{i}:D{i}")

    # Domain assessment
    ws["A13"] = "DOMAIN MATURITY ASSESSMENT"
    ws["A13"].font = Font(bold=True, size=11)

    domain_headers = ["Domain", "Current Level", "Target Level", "Gap", "Priority", "Key Actions", "Notes"]
    for col, header in enumerate(domain_headers, start=1):
        cell = ws.cell(row=14, column=col, value=header)
        apply_style(cell, styles["column_header"])

    domains = [
        ["Policy & Governance", "", "", "", "", "", ""],
        ["IP Identification", "", "", "", "", "", ""],
        ["IP Protection Controls", "", "", "", "", "", ""],
        ["Third-Party Compliance", "", "", "", "", "", ""],
        ["Records Classification", "", "", "", "", "", ""],
        ["Records Protection", "", "", "", "", "", ""],
        ["Retention Management", "", "", "", "", "", ""],
        ["Disposal Process", "", "", "", "", "", ""],
        ["Monitoring & Metrics", "", "", "", "", "", ""],
        ["OVERALL MATURITY", "", "", "", "", "", ""],
    ]

    for row_idx, domain in enumerate(domains, start=15):
        for col_idx, value in enumerate(domain, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            if col_idx > 1:
                cell.fill = styles["input_cell"]["fill"]
        if domain[0] == "OVERALL MATURITY":
            ws.cell(row=row_idx, column=1).font = Font(bold=True)

    dv_level = DataValidation(
        type="list",
        formula1='"5 - Optimised,4 - Managed,3 - Defined,2 - Developing,1 - Initial,0 - Non-existent"',
        allow_blank=True
    )
    ws.add_data_validation(dv_level)
    dv_level.add("B15:C24")

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_priority)
    dv_priority.add("E15:E24")

    set_column_widths(ws, {
        "A": 25, "B": 18, "C": 18, "D": 8, "E": 12, "F": 35, "G": 25
    })
    logger.info("Created Maturity_Assessment sheet")


def create_risk_register_sheet(ws, styles):
    """Create the Risk Register sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "Information Protection Risk Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = "Risks related to IP protection, records protection, and retention"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Risk ID", "Risk Description", "Risk Category", "Likelihood",
        "Impact", "Risk Score", "Current Mitigation", "Residual Risk",
        "Owner", "Status", "Review Date"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Sample risks
    sample_risks = [
        ["RSK-IP-001", "Trade secret exposure through departing employee",
         "IP Protection", "", "", "", "NDA, exit interview, access revocation",
         "", "CISO", "", ""],
        ["RSK-IP-002", "Software license audit findings",
         "IP Protection", "", "", "", "SAM tool, quarterly reconciliation",
         "", "IT Ops", "", ""],
        ["RSK-REC-001", "Records destroyed before retention expires",
         "Records Protection", "", "", "", "Disposal approval workflow",
         "", "Records Mgr", "", ""],
        ["RSK-REC-002", "Records integrity compromised",
         "Records Protection", "", "", "", "Checksums, audit logging",
         "", "CISO", "", ""],
    ]

    for row_idx, risk in enumerate(sample_risks, start=5):
        for col_idx, value in enumerate(risk, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx in [4, 5, 6, 8, 10, 11]:
                cell.fill = styles["input_cell"]["fill"]

    # Add empty rows
    for row in range(9, 25):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_category = DataValidation(
        type="list",
        formula1='"IP Protection,Records Protection,Compliance"',
        allow_blank=True
    )
    ws.add_data_validation(dv_category)
    dv_category.add("C5:C30")

    dv_likelihood = DataValidation(
        type="list",
        formula1='"5 - Almost Certain,4 - Likely,3 - Possible,2 - Unlikely,1 - Rare"',
        allow_blank=True
    )
    ws.add_data_validation(dv_likelihood)
    dv_likelihood.add("D5:D30")

    dv_impact = DataValidation(
        type="list",
        formula1='"5 - Catastrophic,4 - Major,3 - Moderate,2 - Minor,1 - Insignificant"',
        allow_blank=True
    )
    ws.add_data_validation(dv_impact)
    dv_impact.add("E5:E30")

    dv_residual = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_residual)
    dv_residual.add("H5:H30")

    dv_status = DataValidation(
        type="list",
        formula1='"Open,Mitigated,Accepted,Transferred"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("J5:J30")

    set_column_widths(ws, {
        "A": 12, "B": 40, "C": 18, "D": 18, "E": 18,
        "F": 12, "G": 35, "H": 12, "I": 15, "J": 12, "K": 12
    })
    logger.info("Created Risk_Register sheet")


def create_remediation_tracker_sheet(ws, styles):
    """Create the Remediation Tracker sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Remediation Action Tracker"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Track remediation actions from all A.5.32-33 assessments"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Action ID", "Source", "Description", "Priority",
        "Owner", "Due Date", "Progress", "Status", "Blocker", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty rows
    for row in range(5, 40):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_source = DataValidation(
        type="list",
        formula1='"A.5.32-33.1,A.5.32-33.2,A.5.32-33.3,Risk Assessment,Audit Finding"',
        allow_blank=True
    )
    ws.add_data_validation(dv_source)
    dv_source.add("B5:B50")

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_priority)
    dv_priority.add("D5:D50")

    dv_progress = DataValidation(
        type="list",
        formula1='"0%,25%,50%,75%,100%"',
        allow_blank=True
    )
    ws.add_data_validation(dv_progress)
    dv_progress.add("G5:G50")

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Complete,Overdue,On Hold"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H50")

    set_column_widths(ws, {
        "A": 12, "B": 15, "C": 45, "D": 12, "E": 15,
        "F": 12, "G": 10, "H": 12, "I": 25, "J": 25
    })
    logger.info("Created Remediation_Tracker sheet")


def create_trend_analysis_sheet(ws, styles):
    """Create the Trend Analysis sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Trend Analysis"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = "Period-over-period comparison of key metrics"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Metric", "Q1", "Q2", "Q3", "Q4", "YoY Change", "Trend", "Target", "On Track"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Sample metrics for trend tracking
    metrics = [
        "IP Assets with Owners",
        "Third-Party IP Compliance",
        "Software License Compliance",
        "Records Classification Coverage",
        "Protection Control Effectiveness",
        "On-Schedule Disposal Rate",
        "Destruction Certificate Rate",
        "Open Remediation Actions",
        "High-Risk Issues",
        "Overall Maturity Score",
    ]

    for row_idx, metric in enumerate(metrics, start=5):
        ws.cell(row=row_idx, column=1, value=metric).border = styles["border"]
        for col in range(2, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_trend = DataValidation(
        type="list",
        formula1='"Improving,Stable,Declining"',
        allow_blank=True
    )
    ws.add_data_validation(dv_trend)
    dv_trend.add("G5:G20")

    dv_track = DataValidation(
        type="list",
        formula1='"Yes,No,At Risk"',
        allow_blank=True
    )
    ws.add_data_validation(dv_track)
    dv_track.add("I5:I20")

    set_column_widths(ws, {
        "A": 30, "B": 10, "C": 10, "D": 10, "E": 10,
        "F": 12, "G": 12, "H": 10, "I": 10
    })
    logger.info("Created Trend_Analysis sheet")


def create_evidence_register_sheet(ws, styles):
    """Create the Evidence Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Evidence Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Audit evidence supporting dashboard metrics"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Evidence ID", "Description", "Evidence Type",
        "Related Item", "Storage Location", "Collected Date",
        "Collected By", "Verification Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty rows
    for row in range(5, 35):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_type = DataValidation(
        type="list",
        formula1='"Assessment,Report,Metrics,Approval,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C5:C50")

    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Pending,Expired"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H50")

    set_column_widths(ws, {
        "A": 12, "B": 40, "C": 15, "D": 20,
        "E": 35, "F": 15, "G": 20, "H": 18
    })
    logger.info("Created Evidence_Register sheet")


def create_approval_sheet(ws, styles):
    """Create the Approval Sign-Off sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "Dashboard Review and Approval"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "Formal approval of compliance dashboard"
    apply_style(ws["A2"], styles["subheader"])

    # Dashboard summary
    ws["A4"] = "Review Period:"
    ws["B4"] = ""
    ws["A5"] = "Overall Compliance:"
    ws["B5"] = ""
    ws["A6"] = "Maturity Level:"
    ws["B6"] = ""
    ws["A7"] = "Open Risks:"
    ws["B7"] = ""
    ws["A8"] = "Open Remediation Actions:"
    ws["B8"] = ""

    for row in range(4, 9):
        ws[f"A{row}"].font = Font(bold=True)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]

    # Approval table
    ws["A11"] = "APPROVAL SIGNATURES"
    ws["A11"].font = Font(bold=True, size=12)

    approval_headers = ["Role", "Name", "Signature", "Date", "Decision", "Comments"]
    for col, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=13, column=col, value=header)
        apply_style(cell, styles["column_header"])

    approvers = [
        "Chief Information Security Officer",
        "Legal Counsel",
        "Records Manager",
        "Compliance Officer",
        "Internal Audit Representative",
        "Executive Management Representative",
    ]

    for row_idx, role in enumerate(approvers, start=14):
        ws.cell(row=row_idx, column=1, value=role).border = styles["border"]
        for col in range(2, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected,Pending"',
        allow_blank=True
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add("E14:E22")

    set_column_widths(ws, {"A": 35, "B": 25, "C": 20, "D": 15, "E": 22, "F": 30})
    logger.info("Created Approval_SignOff sheet")


# =============================================================================
# MAIN FUNCTION
# =============================================================================
def main():
    """Main entry point."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS-IMP-A.5.32-33.4 Compliance Dashboard Generator")
        logger.info("=" * 70)

        wb = Workbook()
        styles = setup_styles()

        ws_instructions = wb.active
        ws_instructions.title = "Instructions"

        ws_exec = wb.create_sheet("Executive_Summary")
        ws_metrics = wb.create_sheet("Compliance_Metrics")
        ws_control = wb.create_sheet("Control_Assessment")
        ws_maturity = wb.create_sheet("Maturity_Assessment")
        ws_risk = wb.create_sheet("Risk_Register")
        ws_remediation = wb.create_sheet("Remediation_Tracker")
        ws_trend = wb.create_sheet("Trend_Analysis")
        ws_evidence = wb.create_sheet("Evidence_Register")
        ws_approval = wb.create_sheet("Approval_SignOff")

        create_instructions_sheet(ws_instructions, styles)
        create_executive_summary_sheet(ws_exec, styles)
        create_compliance_metrics_sheet(ws_metrics, styles)
        create_control_assessment_sheet(ws_control, styles)
        create_maturity_assessment_sheet(ws_maturity, styles)
        create_risk_register_sheet(ws_risk, styles)
        create_remediation_tracker_sheet(ws_remediation, styles)
        create_trend_analysis_sheet(ws_trend, styles)
        create_evidence_register_sheet(ws_evidence, styles)
        create_approval_sheet(ws_approval, styles)

        wb.save(OUTPUT_FILENAME)
        logger.info("=" * 70)
        logger.info("SUCCESS: Workbook saved as %s", OUTPUT_FILENAME)
        logger.info("=" * 70)
        return 0

    except Exception as e:
        logger.error("FAILED: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.32-33 Information Protection
# =============================================================================
