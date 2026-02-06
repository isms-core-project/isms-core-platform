#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.32-33.S1 - IP Rights Inventory and Compliance Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.32: Intellectual Property Rights
"The organisation should implement appropriate procedures to protect
intellectual property rights."

Assessment Domain 1 of 4: IP Rights Inventory and Compliance

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel workbook for documenting and
managing the organisation's intellectual property assets and third-party
IP compliance.

**Purpose:**
Establish a complete inventory of intellectual property assets, assess
protection controls, and verify compliance with third-party IP rights.

**Generated Workbook Structure:**
1. Instructions - Guidance on IP inventory completion
2. IP_Asset_Inventory - Core register of all organisational IP
3. IP_Protection_Assessment - Protection controls per asset
4. Third_Party_IP_Register - Licensed software and content
5. Software_License_Compliance - License reconciliation
6. Gap_Analysis - Identified gaps and remediation
7. Evidence_Register - Audit evidence tracking
8. Approval_SignOff - Assessment approval

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.32-33.S1"
WORKBOOK_NAME = "IP Rights Inventory"
CONTROL_ID = "A.5.32"
CONTROL_NAME = "Intellectual Property Rights"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "high_value": {
            "fill": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            "font": Font(bold=True, color="FFFFFF"),
        },
        "medium_value": {
            "fill": PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid"),
            "font": Font(bold=True),
        },
        "low_value": {
            "fill": PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid"),
            "font": Font(bold=True),
        },
        "compliant": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "warning": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "non_compliant": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    for attr, value in style_dict.items():
        if attr != "border":
            setattr(cell, attr, value)


def set_column_widths(ws, widths):
    """Set column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    apply_style(ws["A2"], styles["subheader"])

    instructions = [
        "",
        "PURPOSE",
        "This workbook documents the organisation's intellectual property assets and verifies",
        "compliance with third-party IP rights as required by ISO 27001:2022 Control A.5.32.",
        "",
        "IP CATEGORIES",
        "- Trade Secrets: Confidential business information providing competitive advantage",
        "- Patents: Inventions with formal legal protection (granted or pending)",
        "- Copyrights: Original works of authorship (software, documentation, content)",
        "- Trademarks: Brand identifiers (logos, names, slogans)",
        "",
        "PROTECTION REQUIREMENTS BY CATEGORY",
        "Trade Secrets: Access control, NDAs, encryption, documentation of reasonable measures",
        "Patents: Registration tracking, maintenance fee payment, geographic coverage",
        "Copyrights: Notices on works, registration (if applicable), license management",
        "Trademarks: Registration, renewal tracking, brand usage guidelines",
        "",
        "HOW TO USE THIS WORKBOOK",
        "1. Complete IP_Asset_Inventory with all organisational IP",
        "2. Document protection controls in IP_Protection_Assessment",
        "3. Register all third-party IP in Third_Party_IP_Register",
        "4. Reconcile software licenses in Software_License_Compliance",
        "5. Document gaps in Gap_Analysis",
        "6. Collect evidence in Evidence_Register",
        "7. Obtain approval in Approval_SignOff",
        "",
        "KEY STAKEHOLDERS",
        "- Legal Counsel: IP classification and legal protection",
        "- CISO: Technical protection controls",
        "- IT Operations: Software asset management",
        "- Business Unit Leaders: IP identification",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, line in enumerate(instructions, start=4):
        ws[f"A{i}"] = line
        if line in ["PURPOSE", "IP CATEGORIES", "PROTECTION REQUIREMENTS BY CATEGORY",
                    "HOW TO USE THIS WORKBOOK", "KEY STAKEHOLDERS"]:
            ws[f"A{i}"].font = Font(bold=True, size=11)

    set_column_widths(ws, {"A": 100})
    logger.info("Created Instructions sheet")


def create_ip_asset_inventory_sheet(ws, styles):
    """Create the IP Asset Inventory sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "IP Asset Inventory"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:L2")
    ws["A2"] = "Comprehensive register of all organisational intellectual property"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "IP Asset ID", "IP Asset Name", "IP Category", "Description",
        "IP Owner", "Custodian", "Legal Protection Status", "Business Value",
        "Classification", "Creation Date", "Last Review", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Sample data rows
    sample_data = [
        ["IP-001", "Customer Analytics Algorithm", "Trade Secret",
         "Proprietary algorithm for customer churn prediction",
         "Chief Data Officer", "Analytics Team Lead", "Unregistered", "High",
         "Confidential", "", "", ""],
        ["IP-002", "Company Logo", "Trademark",
         "Primary corporate logo and brand identity",
         "Marketing Director", "Brand Manager", "Registered", "High",
         "Internal", "", "", ""],
        ["IP-003", "Product Documentation", "Copyright",
         "User manuals and technical documentation",
         "Product Manager", "Technical Writer Lead", "Unregistered", "Medium",
         "Internal", "", "", ""],
    ]

    for row_idx, data in enumerate(sample_data, start=5):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Add empty input rows
    for row in range(8, 30):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    # Data validations
    dv_category = DataValidation(
        type="list",
        formula1='"Trade Secret,Patent,Copyright,Trademark"',
        allow_blank=True
    )
    ws.add_data_validation(dv_category)
    dv_category.add("C5:C50")

    dv_status = DataValidation(
        type="list",
        formula1='"Registered,Pending,Unregistered,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("G5:G50")

    dv_value = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_value)
    dv_value.add("H5:H50")

    dv_class = DataValidation(
        type="list",
        formula1='"Restricted,Confidential,Internal,Public"',
        allow_blank=True
    )
    ws.add_data_validation(dv_class)
    dv_class.add("I5:I50")

    set_column_widths(ws, {
        "A": 12, "B": 30, "C": 15, "D": 45,
        "E": 20, "F": 20, "G": 18, "H": 12,
        "I": 15, "J": 12, "K": 12, "L": 30
    })
    logger.info("Created IP_Asset_Inventory sheet")


def create_ip_protection_assessment_sheet(ws, styles):
    """Create the IP Protection Assessment sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "IP Protection Assessment"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = "Protection controls assessment for each IP asset"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "IP Asset ID", "IP Asset Name", "Access Control", "Technical Controls",
        "Administrative Controls", "Physical Controls", "Legal Protection",
        "Control Effectiveness", "Gap Description", "Remediation Needed", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 30):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_effectiveness = DataValidation(
        type="list",
        formula1='"Effective,Partial,Ineffective"',
        allow_blank=True
    )
    ws.add_data_validation(dv_effectiveness)
    dv_effectiveness.add("H5:H50")

    dv_status = DataValidation(
        type="list",
        formula1='"Complete,In Progress,Not Started"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("K5:K50")

    set_column_widths(ws, {
        "A": 12, "B": 25, "C": 30, "D": 30,
        "E": 30, "F": 25, "G": 25, "H": 18,
        "I": 35, "J": 35, "K": 15
    })
    logger.info("Created IP_Protection_Assessment sheet")


def create_third_party_ip_register_sheet(ws, styles):
    """Create the Third-Party IP Register sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "Third-Party IP Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = "Licensed software, content, and third-party intellectual property"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Third-Party IP ID", "Software/Content Name", "Vendor", "License Type",
        "License Quantity", "Deployed Quantity", "Compliance Status",
        "Contract Reference", "Renewal Date", "Open Source License", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Sample data
    sample_data = [
        ["TP-001", "Microsoft 365", "Microsoft", "Subscription",
         "500", "487", "Compliant", "MSA-2024-001", "", "N/A", ""],
        ["TP-002", "Adobe Creative Cloud", "Adobe", "Subscription",
         "50", "52", "Over-deployed", "ADO-2024-015", "", "N/A", "Need 2 additional licenses"],
        ["TP-003", "React Framework", "Meta (Open Source)", "Open Source",
         "N/A", "N/A", "Compliant", "", "", "MIT", "Check attribution requirements"],
    ]

    for row_idx, data in enumerate(sample_data, start=5):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Add empty input rows
    for row in range(8, 50):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_license_type = DataValidation(
        type="list",
        formula1='"Perpetual,Subscription,Open Source,Freeware"',
        allow_blank=True
    )
    ws.add_data_validation(dv_license_type)
    dv_license_type.add("D5:D60")

    dv_compliance = DataValidation(
        type="list",
        formula1='"Compliant,Over-deployed,Under-utilised"',
        allow_blank=True
    )
    ws.add_data_validation(dv_compliance)
    dv_compliance.add("G5:G60")

    dv_oss = DataValidation(
        type="list",
        formula1='"GPL,Apache,MIT,BSD,LGPL,MPL,Other,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_oss)
    dv_oss.add("J5:J60")

    set_column_widths(ws, {
        "A": 15, "B": 30, "C": 20, "D": 15,
        "E": 15, "F": 15, "G": 18, "H": 20,
        "I": 12, "J": 18, "K": 30
    })
    logger.info("Created Third_Party_IP_Register sheet")


def create_software_license_compliance_sheet(ws, styles):
    """Create the Software License Compliance sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Software License Compliance"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "License entitlement vs deployment reconciliation"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Software Name", "Vendor", "License Model", "Entitled",
        "Deployed", "Variance", "Compliance Risk", "Remediation Action",
        "Due Date", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows with formula for variance
    for row in range(5, 40):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]
        # Variance formula
        ws.cell(row=row, column=6).value = f"=IF(AND(D{row}<>\"\",E{row}<>\"\"),E{row}-D{row},\"\")"

    dv_model = DataValidation(
        type="list",
        formula1='"Named User,Device,Enterprise,Per Core,Subscription"',
        allow_blank=True
    )
    ws.add_data_validation(dv_model)
    dv_model.add("C5:C50")

    dv_risk = DataValidation(
        type="list",
        formula1='"High,Medium,Low,None"',
        allow_blank=True
    )
    ws.add_data_validation(dv_risk)
    dv_risk.add("G5:G50")

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Complete,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("J5:J50")

    set_column_widths(ws, {
        "A": 30, "B": 20, "C": 18, "D": 12,
        "E": 12, "F": 12, "G": 15, "H": 35,
        "I": 12, "J": 15
    })
    logger.info("Created Software_License_Compliance sheet")


def create_gap_analysis_sheet(ws, styles):
    """Create the Gap Analysis sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Gap Analysis"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Identified protection and compliance gaps with remediation tracking"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Gap ID", "Gap Category", "Description", "Related IP",
        "Risk Rating", "Remediation Action", "Owner",
        "Due Date", "Status", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 35):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_category = DataValidation(
        type="list",
        formula1='"Protection,Compliance,Documentation,Process"',
        allow_blank=True
    )
    ws.add_data_validation(dv_category)
    dv_category.add("B5:B50")

    dv_risk = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_risk)
    dv_risk.add("E5:E50")

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Complete,Accepted"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("I5:I50")

    set_column_widths(ws, {
        "A": 10, "B": 15, "C": 45, "D": 15,
        "E": 12, "F": 40, "G": 20, "H": 12,
        "I": 15, "J": 30
    })
    logger.info("Created Gap_Analysis sheet")


def create_evidence_register_sheet(ws, styles):
    """Create the Evidence Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Evidence Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Audit evidence supporting IP protection assessment"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Evidence ID", "Description", "Evidence Type",
        "Related Item", "Storage Location", "Collected Date",
        "Collected By", "Verification Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 35):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_type = DataValidation(
        type="list",
        formula1='"Document,Screenshot,Report,Configuration,Certificate,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C5:C50")

    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Pending Review,Not Verified,Expired"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H50")

    set_column_widths(ws, {
        "A": 12, "B": 40, "C": 18, "D": 15,
        "E": 35, "F": 15, "G": 20, "H": 18
    })
    logger.info("Created Evidence_Register sheet")


def create_approval_sheet(ws, styles):
    """Create the Approval Sign-Off sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "IP Rights Inventory Assessment Approval"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "Formal approval of IP protection assessment"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment metadata
    ws["A4"] = "Assessment Period:"
    ws["B4"] = ""
    ws["A5"] = "Overall Compliance Status:"
    ws["B5"] = ""
    ws["A6"] = "IP Assets Documented:"
    ws["B6"] = ""
    ws["A7"] = "Open Gaps:"
    ws["B7"] = ""

    for row in range(4, 8):
        ws[f"A{row}"].font = Font(bold=True)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]

    # Approval table
    ws["A10"] = "APPROVAL SIGNATURES"
    ws["A10"].font = Font(bold=True, size=12)

    approval_headers = ["Role", "Name", "Signature", "Date", "Decision", "Comments"]
    for col, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=12, column=col, value=header)
        apply_style(cell, styles["column_header"])

    approvers = [
        "Legal Counsel",
        "Chief Information Security Officer",
        "Data Protection Officer",
        "Compliance Officer",
        "Business Unit Representative",
    ]

    for row_idx, role in enumerate(approvers, start=13):
        ws.cell(row=row_idx, column=1, value=role).border = styles["border"]
        for col in range(2, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected,Pending"',
        allow_blank=True
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add("E13:E20")

    set_column_widths(ws, {"A": 35, "B": 25, "C": 20, "D": 15, "E": 22, "F": 30})
    logger.info("Created Approval_SignOff sheet")


# =============================================================================
# MAIN FUNCTION
# =============================================================================
def main():
    """Main entry point."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS-IMP-A.5.32-33.S1 IP Rights Inventory Generator")
        logger.info("=" * 70)

        wb = Workbook()
        styles = setup_styles()

        # Rename default sheet
        ws_instructions = wb.active
        ws_instructions.title = "Instructions"

        # Create all sheets
        ws_inventory = wb.create_sheet("IP_Asset_Inventory")
        ws_protection = wb.create_sheet("IP_Protection_Assessment")
        ws_third_party = wb.create_sheet("Third_Party_IP_Register")
        ws_license = wb.create_sheet("Software_License_Compliance")
        ws_gap = wb.create_sheet("Gap_Analysis")
        ws_evidence = wb.create_sheet("Evidence_Register")
        ws_approval = wb.create_sheet("Approval_SignOff")

        # Populate sheets
        create_instructions_sheet(ws_instructions, styles)
        create_ip_asset_inventory_sheet(ws_inventory, styles)
        create_ip_protection_assessment_sheet(ws_protection, styles)
        create_third_party_ip_register_sheet(ws_third_party, styles)
        create_software_license_compliance_sheet(ws_license, styles)
        create_gap_analysis_sheet(ws_gap, styles)
        create_evidence_register_sheet(ws_evidence, styles)
        create_approval_sheet(ws_approval, styles)

        # Save workbook
        wb.save(OUTPUT_FILENAME)
        logger.info("=" * 70)
        logger.info("SUCCESS: Workbook saved as %s", OUTPUT_FILENAME)
        logger.info("=" * 70)
        return 0

    except Exception as e:
        logger.error("FAILED: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.32-33 Information Protection
# =============================================================================
