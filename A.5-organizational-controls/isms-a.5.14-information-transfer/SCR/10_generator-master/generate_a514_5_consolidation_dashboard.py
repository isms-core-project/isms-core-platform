#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.14.5 - Information Transfer Consolidation Dashboard
================================================================================

ISO/IEC 27001:2022 Control A.5.14: Information Transfer
Consolidation Dashboard: Executive Summary & Cross-Domain Compliance View

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a consolidation dashboard that aggregates data from all
four A.5.14 assessment workbooks into a unified executive view.

**Source Assessment Domains:**
1. Domain 1: Transfer Rules and Procedures (A.5.14.1)
2. Domain 2: Channel Security Assessment (A.5.14.2)
3. Domain 3: Transfer Agreements Register (A.5.14.3)
4. Domain 4: Compliance Monitoring Dashboard (A.5.14.4)

================================================================================
"""

import logging
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.14.5"
WORKBOOK_NAME = "Consolidation Dashboard"
CONTROL_ID = "A.5.14"
CONTROL_NAME = "Information Transfer"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
SECTION_FONT = Font(bold=True, size=12, color="1F4E79")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


def apply_header_style(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def create_instructions_sheet(ws):
    ws.title = "Instructions"
    instructions = [
        [f"ISMS-IMP-A.5.14.5 - Information Transfer Consolidation Dashboard"],
        [""],
        ["PURPOSE"],
        ["This dashboard consolidates compliance data from all four A.5.14 assessment domains"],
        ["into a unified executive view for compliance monitoring and audit readiness."],
        [""],
        ["SOURCE WORKBOOKS"],
        ["  • A.5.14.1: Transfer Rules and Procedures"],
        ["  • A.5.14.2: Channel Security Assessment"],
        ["  • A.5.14.3: Transfer Agreements Register"],
        ["  • A.5.14.4: Compliance Monitoring Dashboard"],
        [""],
        ["COMPLIANCE SCORING"],
        ["• Compliant (Green): ≥90% of requirements met"],
        ["• Partially Compliant (Amber): 50-89% of requirements met"],
        ["• Non-Compliant (Red): <50% of requirements met"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]
    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SOURCE WORKBOOKS", "COMPLIANCE SCORING"]:
                cell.font = Font(bold=True, size=11)
    ws.column_dimensions['A'].width = 90


def create_executive_summary_sheet(ws):
    ws.title = "Executive_Summary"
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="A.5.14 Information Transfer - Executive Summary")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=3, column=1, value="Reporting Period:").font = Font(bold=True)
    ws.cell(row=3, column=2).fill = INPUT_FILL
    ws.cell(row=3, column=4, value="Assessment Date:").font = Font(bold=True)
    ws.cell(row=3, column=5).fill = INPUT_FILL

    ws.cell(row=5, column=1, value="OVERALL COMPLIANCE STATUS").font = SECTION_FONT
    headers = ["Domain", "Workbook", "Status", "Score %", "Critical Gaps", "Last Updated"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=header)
    apply_header_style(ws, 6, 1, len(headers))

    domains = [
        ("A.5.14", "Transfer Rules and Procedures", "", "", "", ""),
        ("A.5.14", "Channel Security Assessment", "", "", "", ""),
        ("A.5.14", "Transfer Agreements Register", "", "", "", ""),
        ("A.5.14", "Compliance Monitoring Dashboard", "", "", "", ""),
        ("OVERALL", "Consolidated Assessment", "", "", "", ""),
    ]
    for row_num, domain_data in enumerate(domains, 7):
        for col_num, value in enumerate(domain_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = THIN_BORDER
            if col_num in [3, 4, 5]:
                cell.fill = INPUT_FILL
            if row_num == 11:
                cell.font = Font(bold=True)

    ws.cell(row=13, column=1, value="KEY METRICS").font = SECTION_FONT
    metrics_headers = ["Metric", "Target", "Actual", "Status"]
    for col, header in enumerate(metrics_headers, 1):
        ws.cell(row=14, column=col, value=header)
    apply_header_style(ws, 14, 1, len(metrics_headers))

    metrics = [
        ("Transfer procedures documented", "100%", "", ""),
        ("Channels security assessed", "100%", "", ""),
        ("Third-party agreements current", "100%", "", ""),
        ("Transfer incidents resolved", "100%", "", ""),
    ]
    for row_num, metric_data in enumerate(metrics, 15):
        for col_num, value in enumerate(metric_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = THIN_BORDER
            if col_num in [3, 4]:
                cell.fill = INPUT_FILL

    for col, width in enumerate([35, 35, 18, 12, 15, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_domain_overview_sheet(ws):
    ws.title = "Domain_Overview"
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="Domain-by-Domain Compliance Overview")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    domains = [
        (3, "DOMAIN 1: TRANSFER RULES (A.5.14.1)", [
            "Transfer policy documented", "Electronic transfer rules",
            "Physical transfer rules", "Verbal transfer guidelines"]),
        (10, "DOMAIN 2: CHANNEL SECURITY (A.5.14.2)", [
            "Email security assessed", "Cloud service security",
            "File transfer security", "Physical channel security"]),
        (17, "DOMAIN 3: AGREEMENTS (A.5.14.3)", [
            "Agreement register maintained", "Requirements documented",
            "Third-party assessments current", "Review schedule maintained"]),
    ]

    headers = ["Requirement", "Status", "Evidence Ref", "Gap Description", "Remediation"]
    for start_row, section_title, requirements in domains:
        ws.cell(row=start_row, column=1, value=section_title).font = SECTION_FONT
        for col, header in enumerate(headers, 1):
            ws.cell(row=start_row + 1, column=col, value=header)
        apply_header_style(ws, start_row + 1, 1, len(headers))

        for row_num, req in enumerate(requirements, start_row + 2):
            ws.cell(row=row_num, column=1, value=req).border = THIN_BORDER
            for col in range(2, 6):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = INPUT_FILL
                cell.border = THIN_BORDER

    for col, width in enumerate([40, 18, 15, 35, 35], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_transfer_compliance_sheet(ws):
    ws.title = "Transfer_Compliance"
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="Transfer Methods Compliance - Consolidated View")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["Transfer_Method", "Policy_Defined", "Security_Controls", "Encryption_Status",
               "Monitoring_Active", "Compliance_Status", "Gap_Notes", "Owner"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    apply_header_style(ws, 3, 1, len(headers))

    methods = ["Email (Internal)", "Email (External)", "Cloud Storage", "SFTP/SCP",
               "USB/Removable Media", "Courier/Physical", "API Integration", "Verbal/Telephone"]
    for row_num, method in enumerate(methods, 4):
        ws.cell(row=row_num, column=1, value=method).border = THIN_BORDER
        for col in range(2, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    for col, width in enumerate([22, 16, 18, 18, 16, 18, 30, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_channel_compliance_sheet(ws):
    ws.title = "Channel_Compliance"
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="Channel Security Compliance - Consolidated View")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["Channel_Type", "Service/Platform", "Security_Rating", "Last_Assessed",
               "Risk_Level", "Compliance_Status", "Gap_Notes", "Owner"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    apply_header_style(ws, 3, 1, len(headers))

    channels = ["Email - Microsoft 365", "Email - Gateway", "Cloud - SharePoint",
                "Cloud - OneDrive", "File Transfer - SFTP", "Messaging - Teams",
                "API Gateway", "VPN Tunnel"]
    for row_num, channel in enumerate(channels, 4):
        ws.cell(row=row_num, column=1, value=channel).border = THIN_BORDER
        for col in range(2, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    for col, width in enumerate([22, 22, 16, 15, 12, 18, 30, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_agreements_compliance_sheet(ws):
    ws.title = "Agreements_Compliance"
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="Transfer Agreements Compliance - Consolidated View")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["Third_Party", "Agreement_Type", "Transfer_Type", "Expiry_Date",
               "Review_Status", "Compliance_Status", "Gap_Notes", "Owner"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    apply_header_style(ws, 3, 1, len(headers))

    for row_num in range(4, 14):
        ws.cell(row=row_num, column=1, value=f"Third Party {row_num-3}").border = THIN_BORDER
        for col in range(2, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    for col, width in enumerate([25, 20, 18, 15, 15, 18, 30, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_cross_domain_gaps_sheet(ws):
    ws.title = "Cross_Domain_Gaps"
    ws.merge_cells('A1:J1')
    title_cell = ws.cell(row=1, column=1, value="Cross-Domain Gap Analysis")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["Gap_ID", "Source_Domain", "Gap_Description", "Risk_Rating", "Priority",
               "Affected_Controls", "Root_Cause", "Remediation_Action", "Owner", "Target_Date"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    apply_header_style(ws, 3, 1, len(headers))

    for row_num in range(4, 19):
        ws.cell(row=row_num, column=1, value=f"GAP-{row_num-3:03d}").border = THIN_BORDER
        for col in range(2, 11):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    for col, width in enumerate([10, 22, 40, 12, 10, 18, 30, 35, 18, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_remediation_tracker_sheet(ws):
    ws.title = "Remediation_Tracker"
    ws.merge_cells('A1:K1')
    title_cell = ws.cell(row=1, column=1, value="Consolidated Remediation Action Tracker")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["Action_ID", "Related_Gap", "Source_Domain", "Action_Description", "Priority",
               "Owner", "Start_Date", "Target_Date", "Status", "Progress_%", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    apply_header_style(ws, 3, 1, len(headers))

    for row_num in range(4, 19):
        ws.cell(row=row_num, column=1, value=f"ACT-{row_num-3:03d}").border = THIN_BORDER
        for col in range(2, 12):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    for col, width in enumerate([10, 12, 20, 40, 10, 18, 12, 12, 12, 10, 30], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_kpi_summary_sheet(ws):
    ws.title = "KPI_Summary"
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="Key Performance Indicators - Summary")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["KPI", "Target", "Current", "Previous", "Trend", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    apply_header_style(ws, 3, 1, len(headers))

    kpis = [
        ("% of transfer methods with policies", "100%", "", "", "", ""),
        ("% of channels security assessed", "100%", "", "", "", ""),
        ("% of agreements current (<12 months)", "100%", "", "", "", ""),
        ("Transfer incidents (monthly)", "0", "", "", "", ""),
        ("Mean time to resolve incidents", "<24hrs", "", "", "", ""),
        ("Third-party compliance rate", "≥95%", "", "", "", ""),
    ]
    for row_num, kpi_data in enumerate(kpis, 4):
        for col_num, value in enumerate(kpi_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = THIN_BORDER
            if col_num > 2:
                cell.fill = INPUT_FILL

    for col, width in enumerate([42, 15, 12, 12, 10, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_evidence_index_sheet(ws):
    ws.title = "Evidence_Index"
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="Evidence Index - Cross-Reference to Source Assessments")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["Evidence_ID", "Source_Workbook", "Source_Sheet", "Evidence_Type",
               "Evidence_Description", "Location/Reference", "Date_Captured", "Validation_Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    apply_header_style(ws, 3, 1, len(headers))

    for row_num in range(4, 19):
        ws.cell(row=row_num, column=1, value=f"EV-{row_num-3:03d}").border = THIN_BORDER
        for col in range(2, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    for col, width in enumerate([12, 25, 22, 18, 40, 30, 15, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_trend_dashboard_sheet(ws):
    ws.title = "Trend_Dashboard"
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="Historical Compliance Trend Dashboard")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["Period", "Transfer Rules %", "Channels %", "Agreements %",
               "Overall %", "Incidents", "Remediation Rate", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    apply_header_style(ws, 3, 1, len(headers))

    periods = ["Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025", "Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"]
    for row_num, period in enumerate(periods, 4):
        ws.cell(row=row_num, column=1, value=period).border = THIN_BORDER
        for col in range(2, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    for col, width in enumerate([12, 16, 14, 14, 12, 12, 18, 35], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_approval_signoff_sheet(ws):
    ws.title = "Approval_SignOff"
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="Consolidation Dashboard Approval & Sign-Off")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=3, column=1, value="Assessment Period:").font = Font(bold=True)
    ws.cell(row=3, column=2).fill = INPUT_FILL
    ws.cell(row=4, column=1, value="Consolidation Date:").font = Font(bold=True)
    ws.cell(row=4, column=2).fill = INPUT_FILL

    ws.cell(row=6, column=1, value="APPROVAL WORKFLOW").font = SECTION_FONT
    headers = ["Role", "Name", "Date", "Signature", "Comments"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=7, column=col, value=header)
    apply_header_style(ws, 7, 1, len(headers))

    roles = ["Prepared By", "Reviewed By", "Approved By (CISO)", "Data Owner Sign-Off"]
    for row_num, role in enumerate(roles, 8):
        ws.cell(row=row_num, column=1, value=role).border = THIN_BORDER
        for col in range(2, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    for col, width in enumerate([30, 25, 15, 20, 40], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def generate_workbook():
    logger.info("=" * 70)
    logger.info(f"Generating: {DOCUMENT_ID} - {WORKBOOK_NAME}")
    logger.info(f"Control: {CONTROL_REF}")
    logger.info("=" * 70)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        ("Instructions", create_instructions_sheet),
        ("Executive_Summary", create_executive_summary_sheet),
        ("Domain_Overview", create_domain_overview_sheet),
        ("Transfer_Compliance", create_transfer_compliance_sheet),
        ("Channel_Compliance", create_channel_compliance_sheet),
        ("Agreements_Compliance", create_agreements_compliance_sheet),
        ("Cross_Domain_Gaps", create_cross_domain_gaps_sheet),
        ("Remediation_Tracker", create_remediation_tracker_sheet),
        ("KPI_Summary", create_kpi_summary_sheet),
        ("Evidence_Index", create_evidence_index_sheet),
        ("Trend_Dashboard", create_trend_dashboard_sheet),
        ("Approval_SignOff", create_approval_signoff_sheet),
    ]

    for sheet_name, create_func in sheets:
        logger.info(f"Creating {sheet_name} sheet...")
        create_func(wb.create_sheet())

    wb.save(OUTPUT_FILENAME)
    logger.info("=" * 70)
    logger.info(f"SUCCESS: Workbook saved as {OUTPUT_FILENAME}")
    logger.info(f"Sheets created: {len(wb.sheetnames)}")
    logger.info("=" * 70)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(generate_workbook())
    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        sys.exit(1)

# =============================================================================
# QA_VERIFIED: 2026-02-04
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# =============================================================================
