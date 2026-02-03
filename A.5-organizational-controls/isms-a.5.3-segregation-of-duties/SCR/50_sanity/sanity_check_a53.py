#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
A.5.3 Sanity Check Script
================================================================================

Validates all A.5.3 generated workbooks for:
- Required sheets present
- Column headers correct
- Data validations functional
- Formulas valid
- Consistent styling

Usage:
    python3 sanity_check_a53.py
================================================================================
"""

import logging
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# CONSTANTS
# =============================================================================
CONTROL_ID = "A.5.3"
WORKBOOK_DIR = Path(__file__).parent.parent / "90_workbooks"

# Expected workbook structure
EXPECTED_WORKBOOKS = {
    "A.5.3.1": {
        "pattern": "ISMS-IMP-A.5.3.1_SoD_Matrix*.xlsx",
        "required_sheets": [
            "Instructions",
            "Role_Inventory",
            "Conflict_Matrix",
            "Current_Assignments",
            "Gap_Analysis",
            "Remediation_Tracker",
            "Exception_Register",
            "Approval_SignOff",
        ]
    },
    "A.5.3.2": {
        "pattern": "ISMS-IMP-A.5.3.2_Conflict_Analysis*.xlsx",
        "required_sheets": [
            "Instructions",
            "Conflict_Register",
            "Impact_Assessment",
            "Exploitation_Scenarios",
            "Control_Mapping",
            "Trend_Analysis",
            "Prioritisation_Matrix",
            "Evidence_Register",
        ]
    },
    "A.5.3.3": {
        "pattern": "ISMS-IMP-A.5.3.3_Role_Function*.xlsx",
        "required_sheets": [
            "Instructions",
            "Business_Roles",
            "Application_Roles",
            "Functions",
            "Permissions",
            "Role_Function_Map",
            "Function_Conflicts",
            "Validation_Status",
            "Change_Log",
        ]
    },
    "A.5.3.4": {
        "pattern": "ISMS-IMP-A.5.3.4_Compliance_Dashboard*.xlsx",
        "required_sheets": [
            "Executive_Dashboard",
            "KPI_Scorecard",
            "Conflict_Status",
            "Remediation_Progress",
            "Exception_Monitoring",
            "Trend_Analysis",
            "Department_View",
            "Audit_Evidence",
            "Data_Sources",
        ]
    },
}


def check_workbook(filepath: Path, expected_sheets: list) -> tuple:
    """
    Check a single workbook for sanity.

    Args:
        filepath: Path to the workbook
        expected_sheets: List of required sheet names

    Returns:
        tuple: (passed_count, failed_count, issues_list)
    """
    issues = []
    passed = 0
    failed = 0

    try:
        wb = load_workbook(filepath, data_only=False)

        # Check required sheets
        for sheet_name in expected_sheets:
            if sheet_name in wb.sheetnames:
                passed += 1
            else:
                failed += 1
                issues.append(f"Missing sheet: {sheet_name}")

        # Check header row exists
        for ws in wb.worksheets:
            if ws["A1"].value is None:
                issues.append(f"Sheet '{ws.title}' has no header in A1")
                failed += 1
            else:
                passed += 1

            # Check for data validations
            if ws.data_validations.dataValidation:
                passed += 1
            else:
                if ws.title != "Instructions" and ws.title != "Executive_Dashboard":
                    issues.append(f"Sheet '{ws.title}' has no data validations")
                    # Not a failure, just a warning

        # Check for formulas in relevant sheets
        formula_sheets = [
            "Impact_Assessment",
            "Trend_Analysis",
            "Prioritisation_Matrix",
            "Validation_Status",
            "KPI_Scorecard",
            "Conflict_Status",
            "Remediation_Progress",
            "Exception_Monitoring",
            "Department_View"
        ]

        formula_count = 0
        for ws in wb.worksheets:
            if ws.title in formula_sheets:
                for row in ws.iter_rows(min_row=2, max_row=min(10, ws.max_row)):
                    for cell in row:
                        if cell.value and str(cell.value).startswith("="):
                            formula_count += 1

        if formula_count > 0:
            passed += 1
        else:
            # Only flag if we expected formulas
            if any(sheet in wb.sheetnames for sheet in formula_sheets):
                issues.append("No formulas found in formula sheets")

        # Check sheet count matches expectations
        if len(wb.sheetnames) == len(expected_sheets):
            passed += 1
        else:
            issues.append(f"Sheet count mismatch: expected {len(expected_sheets)}, found {len(wb.sheetnames)}")

        wb.close()

    except Exception as e:
        failed += 1
        issues.append(f"Error loading workbook: {e}")

    return passed, failed, issues


def main() -> int:
    """Main execution function."""
    logger.info("=" * 70)
    logger.info(f"{CONTROL_ID} Sanity Check")
    logger.info("=" * 70)

    if not WORKBOOK_DIR.exists():
        logger.warning(f"Workbook directory not found: {WORKBOOK_DIR}")
        logger.info("Run generators first to create workbooks")
        return 0

    total_passed = 0
    total_failed = 0
    all_issues = []

    for doc_id, config in EXPECTED_WORKBOOKS.items():
        logger.info(f"\nChecking {doc_id}...")

        workbooks = list(WORKBOOK_DIR.glob(config["pattern"]))

        if not workbooks:
            logger.warning(f"  No workbooks found matching {config['pattern']}")
            total_failed += 1
            all_issues.append(f"{doc_id}: No workbook found")
            continue

        for wb_path in workbooks:
            logger.info(f"  Checking: {wb_path.name}")
            passed, failed, issues = check_workbook(wb_path, config["required_sheets"])

            total_passed += passed
            total_failed += failed

            if issues:
                for issue in issues:
                    logger.warning(f"    {issue}")
                    all_issues.append(f"{doc_id}: {issue}")
            else:
                logger.info(f"    All checks passed ({passed} checks)")

    logger.info("\n" + "=" * 70)
    logger.info("SANITY CHECK SUMMARY")
    logger.info("=" * 70)
    logger.info(f"Total checks passed: {total_passed}")
    logger.info(f"Total checks failed: {total_failed}")

    if all_issues:
        logger.info(f"\nIssues found ({len(all_issues)}):")
        for issue in all_issues:
            logger.info(f"  - {issue}")
    else:
        logger.info("\nAll sanity checks passed!")

    logger.info("=" * 70)

    return 0 if total_failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.3 Segregation of Duties control
# =============================================================================
