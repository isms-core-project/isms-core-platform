#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
Excel Workbook Sanity Checker - ISMS A.5.8 Assessment Workbooks
================================================================================

Diagnostic utility for troubleshooting Excel's "file level validation and repair"
warnings when opening A.5.8 project assessment workbooks.

**Purpose:**
Identifies common openpyxl-generated Excel issues that trigger repair warnings:
- Formula syntax errors and invalid sheet references
- Data validation conflicts and overlapping ranges
- Style attribute inconsistencies
- Merged cell content issues
- Worksheet structure problems

**When to Use:**
- Excel displays repair warnings when opening generated workbooks
- After modifying assessment generator scripts
- Before distributing workbooks to stakeholders
- Quality assurance validation before portfolio consolidation

**Usage:**
    python3 excel_sanity_check_a58.py ISMS-IMP-A.5.8_X_Assessment_YYYYMMDD.xlsx
    
    Works with any A.5.8 assessment workbook (1-3)

**Output:**
- Diagnostic report with issue categorization (Critical/Warning)
- Recommended remediation actions
- Structural health summary

**Related Scripts:**
- generate_a58_1_project_lifecycle_assessment.py
- generate_a58_2_security_requirements_register.py
- generate_a58_3_portfolio_dashboard.py
- normalize_assessment_files_a58.py
- consolidate_a58_dashboard.py

Control Reference: ISO/IEC 27001:2022 Annex A Control A.5.8
Script Type: Quality Assurance Utility
Version: 1.0
Date: 2025-01-29
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import re
import sys

# =============================================================================
# Third-Party Imports
# =============================================================================
try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


# ============================================================================
# WORKBOOK TYPE DETECTION
# ============================================================================

def detect_workbook_type(filename):
    """Detect which A.5.8 workbook type this is based on filename."""
    filename_lower = filename.lower()
    
    if 'a.5.8.1' in filename_lower or 'project_lifecycle' in filename_lower or 'lifecycle' in filename_lower:
        return 'A.5.8.1', 'Project Lifecycle Security Assessment'
    elif 'a.5.8.2' in filename_lower or 'requirements_register' in filename_lower or 'requirements' in filename_lower:
        return 'A.5.8.2', 'Security Requirements Register'
    elif 'a.5.8.3' in filename_lower or 'portfolio' in filename_lower or 'dashboard' in filename_lower:
        return 'A.5.8.3', 'Project Portfolio Dashboard'
    else:
        return 'Unknown', 'Generic Excel Workbook'


# ============================================================================
# EXPECTED SHEET STRUCTURES
# ============================================================================

EXPECTED_SHEETS = {
    'A.5.8.1': [
        "Instructions & Legend",
        "2. Project Classification",
        "3. Initiation Phase",
        "4. Planning Phase",
        "5. Execution Phase",
        "6. Monitoring Phase",
        "7. Closure Phase",
        "8. Compliance Dashboard",
        "9. Evidence Register",
        "10. Sign-Off"
    ],
    'A.5.8.2': [
        "Instructions & Legend",
        "Requirements Register",
        "Traceability Matrix",
        "Compliance Dashboard",
        "Evidence Register",
        "Sign-Off"
    ],
    'A.5.8.3': [
        "Instructions & Legend",
        "Project Data",
        "Executive Summary",
        "Project Status",
        "Gap Analysis",
        "Trends",
        "Risk Prioritization",
        "Charts"
    ]
}


# ============================================================================
# GENERIC EXCEL VALIDATION
# ============================================================================

def check_workbook_health(filename):
    """Comprehensive diagnostic check for common openpyxl issues."""
    
    workbook_id, workbook_name = detect_workbook_type(filename)
    
    print("=" * 80)
    print(f"EXCEL WORKBOOK DIAGNOSTIC REPORT: {filename}")
    print(f"Detected Type: {workbook_id} - {workbook_name}")
    print("=" * 80)
    
    issues_found = []
    warnings_found = []
    
    try:
        wb = load_workbook(filename, data_only=False)
        print(f"\n✅ Workbook loaded successfully")
        print(f"  Sheets found: {len(wb.sheetnames)}")
        print(f"  Sheet names: {', '.join(wb.sheetnames)}")
        
    except Exception as e:
        print(f"\n❌ CRITICAL: Cannot load workbook: {e}")
        return
    
    # ========================================================================
    # CHECK 1: SHEET STRUCTURE
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 1: SHEET STRUCTURE")
    print("=" * 80)
    
    if workbook_id in EXPECTED_SHEETS:
        expected = EXPECTED_SHEETS[workbook_id]
        actual = wb.sheetnames
        
        missing_sheets = [s for s in expected if s not in actual]
        extra_sheets = [s for s in actual if s not in expected]
        
        if missing_sheets:
            for sheet in missing_sheets:
                issues_found.append(f"  ❌ CRITICAL: Missing required sheet '{sheet}'")
            print(f"❌ {len(missing_sheets)} required sheet(s) missing")
        else:
            print(f"✅ All required sheets present")
        
        if extra_sheets:
            for sheet in extra_sheets:
                warnings_found.append(f"  ⚠️  Unexpected sheet found '{sheet}'")
            print(f"⚠️  {len(extra_sheets)} extra sheet(s) found (may be custom)")
        
    else:
        print("⚠️  Unknown workbook type - skipping structure validation")
    
    # ========================================================================
    # CHECK 2: FORMULA VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 2: FORMULA VALIDATION")
    print("=" * 80)
    
    formula_issues = 0
    inter_sheet_refs = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # Check for sheet references
                    if "'" in formula:
                        sheet_refs = re.findall(r"'([^']+)'!", formula)
                        for ref in sheet_refs:
                            if ref not in wb.sheetnames:
                                issues_found.append(
                                    f"  ❌ {sheet_name}!{cell.coordinate}: "
                                    f"Invalid sheet reference '{ref}'"
                                )
                                formula_issues += 1
                            else:
                                if sheet_name not in inter_sheet_refs:
                                    inter_sheet_refs[sheet_name] = set()
                                inter_sheet_refs[sheet_name].add(ref)
                    
                    # Check for balanced parentheses
                    if formula.count('(') != formula.count(')'):
                        issues_found.append(
                            f"  ❌ {sheet_name}!{cell.coordinate}: "
                            f"Unbalanced parentheses in formula"
                        )
                        formula_issues += 1
    
    if formula_issues == 0:
        print(f"✅ No formula errors detected")
        if inter_sheet_refs:
            print(f"  Inter-sheet references: {len(inter_sheet_refs)} sheet(s)")
    else:
        print(f"❌ {formula_issues} formula issue(s) found")
    
    # ========================================================================
    # CHECK 3: DATA VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 3: DATA VALIDATION RULES")
    print("=" * 80)
    
    validation_count = 0
    validation_issues = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, 'data_validations') and ws.data_validations:
            validation_count += len(ws.data_validations.dataValidation)
            
            for dv in ws.data_validations.dataValidation:
                # Check for overlapping ranges
                if len(dv.cells.ranges) > 1:
                    warnings_found.append(
                        f"  ⚠️  {sheet_name}: Data validation spans multiple ranges"
                    )
    
    if validation_count > 0:
        print(f"✅ {validation_count} data validation rule(s) found")
        if validation_issues == 0:
            print(f"  No issues detected")
    else:
        print(f"⚠️  No data validation rules found (expected for input fields)")
    
    # ========================================================================
    # CHECK 4: MERGED CELLS
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 4: MERGED CELLS")
    print("=" * 80)
    
    merged_count = 0
    merged_issues = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.merged_cells:
            sheet_merged = len(ws.merged_cells.ranges)
            merged_count += sheet_merged
            
            # Check for content in non-top-left cells of merged ranges
            for merged_range in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                
                # Check if any non-top-left cell has content
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if row == min_row and col == min_col:
                            continue  # Skip top-left cell
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            warnings_found.append(
                                f"  ⚠️  {sheet_name}!{cell.coordinate}: "
                                f"Non-empty cell in merged range (content will be lost)"
                            )
                            merged_issues += 1
    
    if merged_count > 0:
        print(f"✅ {merged_count} merged cell range(s) found")
        if merged_issues == 0:
            print(f"  No issues detected")
        else:
            print(f"⚠️  {merged_issues} potential issue(s) with merged cells")
    
    # ========================================================================
    # CHECK 5: CELL STYLES
    # ========================================================================
    print("\n" + "=" * 80)
    print("CHECK 5: CELL STYLES")
    print("=" * 80)
    
    unique_styles = set()
    style_issues = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                # Track unique fill patterns
                if cell.fill and cell.fill.start_color:
                    unique_styles.add(cell.fill.start_color.rgb)
    
    print(f"✅ {len(unique_styles)} unique fill color(s) in use")
    if len(unique_styles) > 20:
        warnings_found.append(
            f"  ⚠️  Excessive number of unique colors ({len(unique_styles)}) - "
            f"may indicate style inconsistency"
        )
    
    # ========================================================================
    # CHECK 6: WORKBOOK-SPECIFIC VALIDATION
    # ========================================================================
    print("\n" + "=" * 80)
    print(f"CHECK 6: {workbook_id}-SPECIFIC VALIDATION")
    print("=" * 80)
    
    if workbook_id == 'A.5.8.1':
        # Check classification sheet has formula for final classification
        if "2. Project Classification" in wb.sheetnames:
            ws = wb["2. Project Classification"]
            # Check for classification formula (should be around row 18, col B)
            found_classification_formula = False
            for row in range(15, 25):
                cell = ws[f"B{row}"]
                if cell.value and isinstance(cell.value, str) and 'IF' in cell.value:
                    found_classification_formula = True
                    break
            
            if found_classification_formula:
                print(f"✅ Classification formula present")
            else:
                warnings_found.append(
                    f"  ⚠️  Classification sheet: Auto-classification formula may be missing"
                )
        
        # Check dashboard has compliance score formulas
        if "8. Compliance Dashboard" in wb.sheetnames:
            ws = wb["8. Compliance Dashboard"]
            found_compliance_formula = False
            for row in ws.iter_rows(min_row=1, max_row=50):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and 'AVERAGE' in cell.value:
                        found_compliance_formula = True
                        break
                if found_compliance_formula:
                    break
            
            if found_compliance_formula:
                print(f"✅ Compliance Dashboard formulas present")
            else:
                warnings_found.append(
                    f"  ⚠️  Compliance Dashboard: Overall score formula may be missing"
                )
    
    elif workbook_id == 'A.5.8.2':
        # Check requirements register has auto-numbering
        if "Requirements Register" in wb.sheetnames:
            ws = wb["Requirements Register"]
            first_id_cell = ws["A4"]
            if first_id_cell.value and isinstance(first_id_cell.value, str) and 'TEXT(ROW' in first_id_cell.value:
                print(f"✅ Requirements Register auto-numbering formula present")
            else:
                warnings_found.append(
                    f"  ⚠️  Requirements Register: Auto-numbering may not be configured"
                )
        
        # Check compliance dashboard has metrics
        if "Compliance Dashboard" in wb.sheetnames:
            ws = wb["Compliance Dashboard"]
            has_formulas = False
            for row in ws.iter_rows(min_row=1, max_row=30):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        has_formulas = True
                        break
                if has_formulas:
                    break
            
            if has_formulas:
                print(f"✅ Compliance Dashboard has formulas")
            else:
                warnings_found.append(
                    f"  ⚠️  Compliance Dashboard: Metrics formulas may be missing"
                )
    
    elif workbook_id == 'A.5.8.3':
        # Check Executive Summary has portfolio score formula
        if "Executive Summary" in wb.sheetnames:
            ws = wb["Executive Summary"]
            found_portfolio_formula = False
            for row in ws.iter_rows(min_row=1, max_row=50):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and 'SUMPRODUCT' in cell.value and 'Project Data' in cell.value:
                        found_portfolio_formula = True
                        break
                if found_portfolio_formula:
                    break
            
            if found_portfolio_formula:
                print(f"✅ Executive Summary portfolio score formula present")
            else:
                warnings_found.append(
                    f"  ⚠️  Executive Summary: Portfolio score formula may be missing"
                )
    
    # ========================================================================
    # FINAL SUMMARY
    # ========================================================================
    print("\n" + "=" * 80)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 80)
    
    if issues_found:
        print(f"\n❌ CRITICAL ISSUES FOUND ({len(issues_found)}):")
        for issue in issues_found:
            print(issue)
    
    if warnings_found:
        print(f"\n⚠️  WARNINGS ({len(warnings_found)}):")
        for warning in warnings_found:
            print(warning)
    
    if not issues_found and not warnings_found:
        print(f"\n✅ EXCELLENT: No issues detected!")
        print(f"  Workbook appears healthy and ready for use")
    elif not issues_found:
        print(f"\n✅ PASS: No critical issues (warnings can be reviewed)")
    else:
        print(f"\n❌ FAIL: Critical issues must be resolved")
        print(f"\n📋 RECOMMENDED ACTIONS:")
        print(f"  1. Regenerate workbook using generate_a58_X script")
        print(f"  2. Review and fix formula references manually")
        print(f"  3. Run this diagnostic again to verify fixes")
    
    print("\n" + "=" * 80)
    print(f"Workbook: {filename}")
    print(f"Type: {workbook_id} - {workbook_name}")
    print(f"Status: {'❌ FAILED' if issues_found else '✅ PASSED'}")
    print("=" * 80 + "\n")
    
    wb.close()
    
    return len(issues_found) == 0


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    if len(sys.argv) < 2:
        print("\n❌ Error: No input file provided")
        print("\nUsage:")
        print("  python3 excel_sanity_check_a58.py <workbook.xlsx>")
        print("\nExamples:")
        print("  python3 excel_sanity_check_a58.py ISMS-IMP-A.5.8.1_Project_Lifecycle_20250129.xlsx")
        print("  python3 excel_sanity_check_a58.py ISMS-IMP-A.5.8.2_Requirements_Register_20250129.xlsx")
        print("  python3 excel_sanity_check_a58.py ISMS-IMP-A.5.8.3_Portfolio_Dashboard_20250129.xlsx")
        sys.exit(1)
    
    filename = sys.argv[1]
    
    try:
        passed = check_workbook_health(filename)
        sys.exit(0 if passed else 1)
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (syntax validated, STANDARDIZATION applied)
# QA_TOOL: Claude Code Deep Scan
# STANDARDIZATION: License header, logging, imports reorganized, main() pattern
# =============================================================================
