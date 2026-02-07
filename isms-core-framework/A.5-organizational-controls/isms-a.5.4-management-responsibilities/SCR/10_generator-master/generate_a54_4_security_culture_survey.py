#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.4.4 - Security Culture Survey
================================================================================

ISO/IEC 27001:2022 Control A.5.4: Management Responsibilities
Assessment Domain 4 of 4: Security Culture Survey

This script generates an Excel workbook for conducting annual employee security
perception surveys with year-over-year trend analysis and action plan generation.

Per ISMS Copilot Audit Requirements.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.4.4"
WORKBOOK_NAME = "Security Culture Survey"
CONTROL_ID = "A.5.4"
CONTROL_NAME = "Management Responsibilities"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# SURVEY QUESTIONS (Per ISMS Best Practices)
# =============================================================================
SURVEY_CATEGORIES = {
    "Leadership Commitment": [
        "Management demonstrates commitment to information security",
        "Security is a visible priority at the executive level",
        "Leaders allocate adequate resources for security initiatives",
        "Management responds appropriately to security incidents",
    ],
    "Policy Awareness": [
        "I understand the organization's security policies",
        "Security policies are clearly communicated",
        "I know where to find security policies and procedures",
        "Policy updates are communicated effectively",
    ],
    "Training Effectiveness": [
        "Security training is relevant to my role",
        "I feel confident handling security situations after training",
        "Training materials are engaging and useful",
        "Refresher training helps maintain awareness",
    ],
    "Reporting Culture": [
        "I feel comfortable reporting security concerns",
        "Reported issues are addressed in a timely manner",
        "There is no fear of blame when reporting incidents",
        "I know how and where to report security issues",
    ],
    "Personal Responsibility": [
        "I understand my personal security responsibilities",
        "Security is part of my daily work routine",
        "I actively look for security risks in my work",
        "I take ownership of protecting company information",
    ],
}


# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions sheet."""
    ws.title = "Instructions"

    instructions = [
        ["ISMS-IMP-A.5.4.4 - Security Culture Survey"],
        [""],
        ["PURPOSE"],
        ["This workbook supports annual employee security culture assessments"],
        ["per ISMS-POL-A.5.4 (Management Responsibilities)."],
        [""],
        ["SHEETS"],
        ["1. Instructions - This guidance sheet"],
        ["2. Survey Questions - Standard survey questionnaire"],
        ["3. Response Data - Aggregated survey responses by department"],
        ["4. YoY Trend Analysis - Year-over-year comparison"],
        ["5. Action Plan - Improvement actions based on findings"],
        ["6. Executive Summary - High-level results for leadership"],
        [""],
        ["WORKFLOW"],
        ["1. Deploy survey using Survey Questions as template"],
        ["2. Aggregate responses into Response Data sheet"],
        ["3. Compare with prior years in YoY Trend Analysis"],
        ["4. Generate Action Plan for areas scoring below 70%"],
        ["5. Present Executive Summary to management"],
        [""],
        ["SCORING"],
        ["- Response scale: 1 (Strongly Disagree) to 5 (Strongly Agree)"],
        ["- Category scores averaged across all questions"],
        ["- Target: 70% minimum (3.5/5.0), 80% target (4.0/5.0)"],
        ["- Action required for any category below 70%"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SHEETS", "WORKFLOW", "SCORING"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 70


def create_survey_questions_sheet(ws):
    """Create the Survey Questions sheet."""
    ws.title = "Survey Questions"

    headers = ["Question_ID", "Category", "Question_Text", "Response_Scale"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    row = 2
    q_num = 1
    for category, questions in SURVEY_CATEGORIES.items():
        for question in questions:
            ws.cell(row=row, column=1, value=f"Q{q_num:02d}").border = THIN_BORDER
            ws.cell(row=row, column=2, value=category).border = THIN_BORDER
            ws.cell(row=row, column=3, value=question).border = THIN_BORDER
            ws.cell(row=row, column=4, value="1-5 (Strongly Disagree to Strongly Agree)").border = THIN_BORDER
            row += 1
            q_num += 1

    # Column widths
    widths = [12, 22, 55, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_response_data_sheet(ws):
    """Create the Response Data sheet."""
    ws.title = "Response Data"

    headers = ["Department", "Respondents", "Response_Rate"]
    # Add category headers
    for category in SURVEY_CATEGORIES.keys():
        headers.append(f"{category}_Avg")
    headers.append("Overall_Avg")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Sample departments
    departments = ["Executive", "Finance", "IT", "Operations", "HR",
                   "Legal", "Sales", "Marketing", "Engineering", "Support", "TOTAL"]

    row = 2
    for dept in departments:
        ws.cell(row=row, column=1, value=dept).border = THIN_BORDER
        if dept == "TOTAL":
            ws.cell(row=row, column=1).font = Font(bold=True)
        for col in range(2, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL if dept != "TOTAL" else LOCKED_FILL
            cell.border = THIN_BORDER
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    for i in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20


def create_yoy_trend_sheet(ws):
    """Create the YoY Trend Analysis sheet."""
    ws.title = "YoY Trend Analysis"

    # Title
    ws.merge_cells('A1:G1')
    cell = ws.cell(row=1, column=1, value="Security Culture Survey - Year-over-Year Trend Analysis")
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")

    headers = ["Category", "Year-3", "Year-2", "Year-1", "Current", "YoY_Change", "Trend"]

    row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    row = 4
    for category in SURVEY_CATEGORIES.keys():
        ws.cell(row=row, column=1, value=category).border = THIN_BORDER
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        # YoY Change formula
        ws.cell(row=row, column=6, value=f"=IF(AND(D{row}>0,E{row}>0),E{row}-D{row},\"-\")").border = THIN_BORDER
        # Trend
        ws.cell(row=row, column=7).fill = INPUT_FILL
        ws.cell(row=row, column=7).border = THIN_BORDER
        row += 1

    # Overall row
    ws.cell(row=row, column=1, value="OVERALL").border = THIN_BORDER
    ws.cell(row=row, column=1).font = Font(bold=True)
    for col in range(2, 8):
        cell = ws.cell(row=row, column=col)
        cell.fill = LOCKED_FILL
        cell.border = THIN_BORDER

    # Trend validation
    trend_dv = DataValidation(type="list", formula1='"Improving,Stable,Declining"')
    ws.add_data_validation(trend_dv)
    trend_dv.add('G4:G15')

    # Column widths
    widths = [25, 10, 10, 10, 10, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_action_plan_sheet(ws):
    """Create the Action Plan sheet."""
    ws.title = "Action Plan"

    # Title
    ws.merge_cells('A1:H1')
    cell = ws.cell(row=1, column=1, value="Security Culture Improvement Action Plan")
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Note: Actions required for categories scoring below 70% (3.5/5.0)")
    ws.cell(row=2, column=1).font = Font(italic=True)

    headers = ["Action_ID", "Category", "Current_Score", "Gap", "Action_Required",
               "Owner", "Due_Date", "Status"]

    row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate with categories
    row = 5
    action_num = 1
    for category in SURVEY_CATEGORIES.keys():
        ws.cell(row=row, column=1, value=f"SCS-{action_num:02d}").border = THIN_BORDER
        ws.cell(row=row, column=2, value=category).border = THIN_BORDER
        for col in range(3, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        # Gap formula
        ws.cell(row=row, column=4, value=f"=IF(C{row}>0,3.5-C{row},\"-\")")
        row += 1
        action_num += 1

    # Additional rows for custom actions
    for r in range(row, row + 10):
        for col in range(1, 9):
            cell = ws.cell(row=r, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Status validation
    status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,Completed,Deferred"')
    ws.add_data_validation(status_dv)
    status_dv.add('H5:H25')

    # Column widths
    widths = [12, 22, 14, 8, 40, 18, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_executive_summary_sheet(ws):
    """Create the Executive Summary sheet."""
    ws.title = "Executive Summary"

    # Title
    ws.merge_cells('A1:E1')
    cell = ws.cell(row=1, column=1, value="Security Culture Survey - Executive Summary")
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value=f"Survey Period: [Year]")
    ws.cell(row=3, column=1, value=f"Report Date: {GENERATED_DATE}")

    # Key Metrics
    row = 5
    ws.cell(row=row, column=1, value="KEY METRICS").font = Font(bold=True, size=12)

    metrics = [
        ("Overall Culture Score", "[Score]/5.0", "[XX]%"),
        ("Survey Response Rate", "[XX]%", "Target: 80%"),
        ("Categories Meeting Target", "[X]/5", "Target: 100%"),
        ("YoY Improvement", "[+/-X.X]", "Target: Positive"),
    ]

    row = 6
    headers = ["Metric", "Value", "Benchmark"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    row = 7
    for metric, value, benchmark in metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=2, value=value).border = THIN_BORDER
        ws.cell(row=row, column=2).fill = INPUT_FILL
        ws.cell(row=row, column=3, value=benchmark).border = THIN_BORDER
        row += 1

    # Category Summary
    row += 2
    ws.cell(row=row, column=1, value="CATEGORY SCORES").font = Font(bold=True, size=12)

    row += 1
    cat_headers = ["Category", "Score", "Status", "Priority Action"]
    for col, header in enumerate(cat_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    row += 1
    for category in SURVEY_CATEGORIES.keys():
        ws.cell(row=row, column=1, value=category).border = THIN_BORDER
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        row += 1

    # Status validation
    status_dv = DataValidation(type="list", formula1='"Exceeds Target,Meets Target,Below Target,Critical"')
    ws.add_data_validation(status_dv)
    status_dv.add(f'C{row-5}:C{row-1}')

    # Column widths
    widths = [25, 15, 15, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_workbook():
    """Generate the complete survey workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()
    default_sheet = wb.active

    create_instructions_sheet(wb.create_sheet())
    create_survey_questions_sheet(wb.create_sheet())
    create_response_data_sheet(wb.create_sheet())
    create_yoy_trend_sheet(wb.create_sheet())
    create_action_plan_sheet(wb.create_sheet())
    create_executive_summary_sheet(wb.create_sheet())

    wb.remove(default_sheet)

    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-02-02
# QA_STATUS: PASSED - NEW GENERATOR PER AUDIT
# QA_TOOL: Claude Code
# CHANGES: New generator created per ISMS Copilot audit requirement for
#          Security Culture Survey (annual perception assessment, YoY trends,
#          action plan generation)
# =============================================================================
