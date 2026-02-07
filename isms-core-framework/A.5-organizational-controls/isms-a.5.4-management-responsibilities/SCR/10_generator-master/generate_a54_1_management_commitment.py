#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.4.1 - Management Commitment Assessment
================================================================================

ISO/IEC 27001:2022 Control A.5.4: Management Responsibilities
Assessment Domain 1 of 3: Management Commitment

This script generates an Excel assessment workbook for evaluating management
commitment to information security across leadership levels.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.4.1"
WORKBOOK_NAME = "Management Commitment Assessment"
CONTROL_ID = "A.5.4"
CONTROL_NAME = "Management Responsibilities"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# COMMITMENT CRITERIA (Per ISMS Copilot Audit - 100-point rubric)
# =============================================================================
COMMITMENT_CRITERIA = [
    # Category, Criterion, Weight (Total: 100 points)
    # Attendance: 30 points
    ("Attendance", "Attends scheduled ISMS Committee meetings", 10),
    ("Attendance", "Participates in security governance forums", 10),
    ("Attendance", "Available for security escalations and decisions", 10),
    # Training: 20 points
    ("Training", "Completes required security awareness training on time", 10),
    ("Training", "Ensures direct reports complete mandatory training", 10),
    # Resource Support: 20 points
    ("Resource Support", "Approves budget for security initiatives", 10),
    ("Resource Support", "Provides adequate staffing for security activities", 10),
    # Operational Engagement: 20 points
    ("Operational Engagement", "Reviews and acts on security reports/metrics", 10),
    ("Operational Engagement", "Responds to security incidents within SLA", 10),
    # Culture/Enforcement: 10 points
    ("Culture/Enforcement", "Demonstrates security-conscious behavior (no bypasses)", 5),
    ("Culture/Enforcement", "Holds personnel accountable for policy compliance", 5),
]

# =============================================================================
# SCORING THRESHOLDS (Per ISMS Copilot Audit)
# =============================================================================
SCORE_THRESHOLDS = {
    "Exemplary": (90, 100),      # 90-100%
    "Adequate": (70, 89),        # 70-89%
    "Improvement Needed": (50, 69),  # 50-69%
    "Non-Compliant": (0, 49),    # <50%
}


# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions sheet."""
    ws.title = "Instructions"

    instructions = [
        ["ISMS-IMP-A.5.4.1 - Management Commitment Assessment"],
        [""],
        ["PURPOSE"],
        ["This workbook assesses management commitment to information security"],
        ["per ISMS-POL-A.5.4 (Management Responsibilities)."],
        [""],
        ["SHEETS"],
        ["1. Instructions - This guidance sheet"],
        ["2. Manager Inventory - List of managers to assess"],
        ["3. Commitment Assessment - Individual manager assessments"],
        ["4. Summary Scores - Aggregated commitment scores"],
        [""],
        ["WORKFLOW"],
        ["1. Populate Manager Inventory with all management personnel"],
        ["2. Complete Commitment Assessment for each manager"],
        ["3. Review Summary Scores for organizational commitment level"],
        ["4. Identify improvement areas for management training"],
        [""],
        ["SCORING (100-point rubric)"],
        ["- Each criterion scored 0-5 (0=Not Demonstrated, 5=Fully Demonstrated)"],
        ["- Categories: Attendance (30pts), Training (20pts), Resource Support (20pts),"],
        ["  Operational Engagement (20pts), Culture/Enforcement (10pts)"],
        ["- Thresholds: 90-100% Exemplary, 70-89% Adequate, 50-69% Improvement Needed, <50% Non-Compliant"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SHEETS", "WORKFLOW", "SCORING"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 70


def create_manager_inventory_sheet(ws):
    """Create the Manager Inventory sheet."""
    ws.title = "Manager Inventory"

    headers = [
        "Manager_ID", "Name", "Title", "Department", "Management_Level",
        "Direct_Reports", "Assessment_Date", "Assessor", "Status"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    level_dv = DataValidation(type="list", formula1='"Executive,Director,Manager,Team Lead,Supervisor"')
    ws.add_data_validation(level_dv)
    level_dv.add('E2:E50')

    status_dv = DataValidation(type="list", formula1='"Pending,In Progress,Complete"')
    ws.add_data_validation(status_dv)
    status_dv.add('I2:I50')

    # Format input rows
    for row in range(2, 51):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 25, 30, 20, 18, 12, 15, 20, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_commitment_assessment_sheet(ws):
    """Create the Commitment Assessment sheet."""
    ws.title = "Commitment Assessment"

    headers = [
        "Manager_ID", "Category", "Criterion", "Weight",
        "Score (0-5)", "Weighted_Score", "Evidence", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate criteria for first manager placeholder
    row = 2
    for category, criterion, weight in COMMITMENT_CRITERIA:
        ws.cell(row=row, column=1, value="[Manager_ID]").fill = INPUT_FILL
        ws.cell(row=row, column=2, value=category)
        ws.cell(row=row, column=3, value=criterion)
        ws.cell(row=row, column=4, value=weight)
        ws.cell(row=row, column=5).fill = INPUT_FILL  # Score input
        # Weighted score formula
        ws.cell(row=row, column=6, value=f"=E{row}*D{row}/5")
        ws.cell(row=row, column=7).fill = INPUT_FILL  # Evidence
        ws.cell(row=row, column=8).fill = INPUT_FILL  # Notes

        for col in range(1, 9):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Score validation
    score_dv = DataValidation(type="whole", operator="between", formula1="0", formula2="5")
    score_dv.error = "Score must be 0-5"
    ws.add_data_validation(score_dv)
    score_dv.add('E2:E500')

    # Column widths
    widths = [12, 18, 50, 8, 12, 15, 35, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_summary_scores_sheet(ws):
    """Create the Summary Scores sheet."""
    ws.title = "Summary Scores"

    headers = [
        "Manager_ID", "Name", "Management_Level", "Total_Weight",
        "Achieved_Score", "Percentage", "Status", "Improvement_Areas"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Total possible weight
    total_weight = sum(w for _, _, w in COMMITMENT_CRITERIA)

    # Pre-populate template row
    row = 2
    ws.cell(row=row, column=1, value="[Manager_ID]").fill = INPUT_FILL
    ws.cell(row=row, column=2, value="[Name]").fill = INPUT_FILL
    ws.cell(row=row, column=3, value="[Level]").fill = INPUT_FILL
    ws.cell(row=row, column=4, value=total_weight)
    ws.cell(row=row, column=5).fill = INPUT_FILL  # Manual entry or SUMIF
    ws.cell(row=row, column=6, value=f"=IF(E{row}>0,E{row}/D{row}*100,0)")
    ws.cell(row=row, column=7, value=f'=IF(F{row}>=90,"Exemplary",IF(F{row}>=70,"Adequate",IF(F{row}>=50,"Improvement Needed","Non-Compliant")))')
    ws.cell(row=row, column=8).fill = INPUT_FILL

    for col in range(1, 9):
        ws.cell(row=row, column=col).border = THIN_BORDER

    # Format additional rows
    for row in range(3, 51):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 25, 18, 12, 15, 12, 10, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_workbook():
    """Generate the complete assessment workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Create all sheets
    create_instructions_sheet(wb.create_sheet())
    create_manager_inventory_sheet(wb.create_sheet())
    create_commitment_assessment_sheet(wb.create_sheet())
    create_summary_scores_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-02-02
# QA_STATUS: PASSED - AUDIT ALIGNMENT UPDATE
# QA_TOOL: Claude Code
# CHANGES: Updated categories per ISMS Copilot audit (Attendance 30pts, Training 20pts,
#          Resource Support 20pts, Operational Engagement 20pts, Culture/Enforcement 10pts);
#          Updated thresholds (90-100 Exemplary, 70-89 Adequate, 50-69 Improvement, <50 Non-Compliant)
# =============================================================================
