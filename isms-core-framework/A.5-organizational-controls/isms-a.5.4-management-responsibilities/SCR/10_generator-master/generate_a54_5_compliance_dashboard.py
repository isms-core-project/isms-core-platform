#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.4.5 - Management Responsibilities Compliance Dashboard
================================================================================

ISO/IEC 27001:2022 Control A.5.4: Management Responsibilities
Assessment Domain 5 of 5: Compliance Dashboard

This script generates a comprehensive compliance dashboard that consolidates
data from all A.5.4 assessment workbooks into a unified view for compliance
monitoring, audit readiness, and executive reporting.

Source Assessment Domains:
1. A.5.4.1 - Management Commitment Assessment
2. A.5.4.2 - Compliance Oversight Tracker
3. A.5.4.3 - Leadership Dashboard
4. A.5.4.4 - Security Culture Survey

================================================================================
"""

import logging
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.4.5"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.5.4"
CONTROL_NAME = "Management Responsibilities"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

SECTION_FONT = Font(bold=True, size=12, color="1F4E79")

COMPLIANT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
NON_COMPLIANT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
CALC_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions sheet."""
    ws.title = "Instructions"

    instructions = [
        [f"ISMS-IMP-A.5.4.5 - Management Responsibilities Compliance Dashboard"],
        [""],
        ["PURPOSE"],
        ["This dashboard consolidates compliance data from all A.5.4 assessment workbooks"],
        ["into a unified view for compliance monitoring, audit readiness, and executive reporting."],
        [""],
        ["SOURCE WORKBOOKS"],
        ["This dashboard aggregates data from:"],
        ["  • A.5.4.1: Management Commitment Assessment - Individual manager scoring"],
        ["  • A.5.4.2: Compliance Oversight Tracker - Training, violations, access reviews"],
        ["  • A.5.4.3: Leadership Dashboard - Executive metrics by department"],
        ["  • A.5.4.4: Security Culture Survey - Employee perception assessment"],
        [""],
        ["SHEETS IN THIS WORKBOOK"],
        ["1. Instructions - This guidance sheet"],
        ["2. Executive_Summary - High-level compliance status"],
        ["3. Compliance_KPIs - Key performance indicators"],
        ["4. Compliance_Scorecard - Requirement-level assessment"],
        ["5. Gap_Analysis - Identified gaps and remediation"],
        ["6. Audit_Readiness - Evidence checklist for audits"],
        ["7. Trend_Analysis - Historical compliance trends"],
        ["8. Evidence_Register - Audit evidence documentation"],
        ["9. Approval_SignOff - Management approval workflow"],
        [""],
        ["COMPLIANCE SCORING"],
        ["• Compliant (Green): ≥90% of requirements met"],
        ["• Partially Compliant (Amber): 70-89% of requirements met"],
        ["• Non-Compliant (Red): <70% of requirements met"],
        [""],
        ["REFRESH FREQUENCY"],
        ["• Executive Summary: Monthly"],
        ["• KPI Dashboard: Quarterly"],
        ["• Full Assessment: Annual or before audits"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SOURCE WORKBOOKS", "SHEETS IN THIS WORKBOOK",
                          "COMPLIANCE SCORING", "REFRESH FREQUENCY"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 90


def create_executive_summary_sheet(ws):
    """Create the Executive Summary sheet."""
    ws.title = "Executive_Summary"

    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="A.5.4 Management Responsibilities - Executive Summary")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    # Reporting info
    ws.cell(row=3, column=1, value="Reporting Period:").font = Font(bold=True)
    ws.cell(row=3, column=2).fill = INPUT_FILL
    ws.cell(row=3, column=4, value="Assessment Date:").font = Font(bold=True)
    ws.cell(row=3, column=5).fill = INPUT_FILL

    # Overall compliance status
    ws.cell(row=5, column=1, value="OVERALL COMPLIANCE STATUS").font = SECTION_FONT

    headers = ["Assessment Domain", "Workbook", "Status", "Score %", "Key Issues"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    domains = [
        ("Management Commitment", "A.5.4.1", "", "", ""),
        ("Compliance Oversight", "A.5.4.2", "", "", ""),
        ("Leadership Metrics", "A.5.4.3", "", "", ""),
        ("Security Culture", "A.5.4.4", "", "", ""),
        ("OVERALL A.5.4", "Consolidated", "", "", ""),
    ]

    row = 7
    for domain_data in domains:
        for col, value in enumerate(domain_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if col >= 3:
                cell.fill = INPUT_FILL
            if row == 11:  # OVERALL row
                cell.font = Font(bold=True)
        row += 1

    # Data validation for status
    status_dv = DataValidation(type="list", formula1='"Compliant,Partially Compliant,Non-Compliant"')
    ws.add_data_validation(status_dv)
    status_dv.add('C7:C11')

    # Key Metrics Section
    ws.cell(row=13, column=1, value="KEY METRICS SUMMARY").font = SECTION_FONT

    metrics_headers = ["Metric", "Target", "Current", "Status", "Trend"]
    for col, header in enumerate(metrics_headers, 1):
        cell = ws.cell(row=14, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    metrics = [
        ("Managers assessed", "100%", "", "", ""),
        ("Managers meeting commitment target (≥70%)", "100%", "", "", ""),
        ("Training oversight compliance", "≥90%", "", "", ""),
        ("Policy violations addressed within SLA", "100%", "", "", ""),
        ("Access reviews completed", "100%", "", "", ""),
        ("Security culture score", "≥70%", "", "", ""),
    ]

    row = 15
    for metric in metrics:
        for col, value in enumerate(metric, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if col >= 3:
                cell.fill = INPUT_FILL
        row += 1

    # Column widths
    widths = [40, 20, 15, 18, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_compliance_kpis_sheet(ws):
    """Create the Compliance KPIs sheet."""
    ws.title = "Compliance_KPIs"

    headers = [
        "KPI_ID", "KPI_Name", "Description", "Target", "Current_Value",
        "Measurement_Method", "Frequency", "Data_Source", "Owner",
        "Status", "Trend", "Comments"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate KPIs
    kpis = [
        ("A54-KPI-001", "Management Commitment Score", "Average commitment score across all managers",
         "≥70%", "", "Sum of weighted scores / Total possible", "Quarterly",
         "A.5.4.1 Summary Scores", "CISO", "", "", ""),
        ("A54-KPI-002", "Manager Assessment Coverage", "% of managers assessed",
         "100%", "", "Assessed managers / Total managers", "Quarterly",
         "A.5.4.1 Manager Inventory", "HR", "", "", ""),
        ("A54-KPI-003", "Training Oversight Rate", "% of teams with training compliance ≥90%",
         "100%", "", "Compliant teams / Total teams", "Quarterly",
         "A.5.4.2 Training Oversight", "Security Manager", "", "", ""),
        ("A54-KPI-004", "Violation Response Time", "Average days to address violations",
         "≤5 days", "", "Sum of response days / Total violations", "Monthly",
         "A.5.4.2 Policy Violations", "Security Manager", "", "", ""),
        ("A54-KPI-005", "Access Review Completion", "% of required access reviews completed",
         "100%", "", "Completed reviews / Required reviews", "Quarterly",
         "A.5.4.2 Access Reviews", "Security Manager", "", "", ""),
        ("A54-KPI-006", "Security Culture Score", "Overall employee perception score",
         "≥70%", "", "Survey average / 5.0 × 100", "Annual",
         "A.5.4.4 Response Data", "CISO", "", "", ""),
        ("A54-KPI-007", "Escalation Trigger Rate", "% of managers triggering escalation thresholds",
         "<10%", "", "Triggered / Total managers", "Quarterly",
         "A.5.4.2 Escalation Triggers", "CISO", "", "", ""),
    ]

    row = 2
    for kpi in kpis:
        for col, value in enumerate(kpi, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if col in [5, 10, 11, 12]:
                cell.fill = INPUT_FILL
        row += 1

    # Add empty input rows
    for row in range(row, row + 8):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validations
    status_dv = DataValidation(type="list", formula1='"On Target,At Risk,Below Target"')
    ws.add_data_validation(status_dv)
    status_dv.add('J2:J20')

    trend_dv = DataValidation(type="list", formula1='"Improving,Stable,Declining"')
    ws.add_data_validation(trend_dv)
    trend_dv.add('K2:K20')

    freq_dv = DataValidation(type="list", formula1='"Monthly,Quarterly,Semi-annual,Annual"')
    ws.add_data_validation(freq_dv)
    freq_dv.add('G2:G20')

    # Column widths
    widths = [15, 28, 45, 12, 15, 35, 12, 25, 18, 12, 12, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_compliance_scorecard_sheet(ws):
    """Create the Compliance Scorecard sheet."""
    ws.title = "Compliance_Scorecard"

    headers = [
        "Requirement_ID", "Requirement", "Source", "Weight",
        "Evidence_Available", "Implementation_Status", "Score",
        "Max_Score", "Gap_Description", "Remediation_Status"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # A.5.4 requirements based on ISO 27001:2022
    requirements = [
        ("REQ-001", "Management requires personnel to apply security per policies", "A.5.4", "15", "", "", "", "15", "", ""),
        ("REQ-002", "All managers assessed for security commitment", "A.5.4.1", "10", "", "", "", "10", "", ""),
        ("REQ-003", "Manager commitment scores meet minimum threshold", "A.5.4.1", "15", "", "", "", "15", "", ""),
        ("REQ-004", "Training oversight tracked and reported", "A.5.4.2", "10", "", "", "", "10", "", ""),
        ("REQ-005", "Policy violations addressed per escalation triggers", "A.5.4.2", "10", "", "", "", "10", "", ""),
        ("REQ-006", "Access reviews completed by managers", "A.5.4.2", "10", "", "", "", "10", "", ""),
        ("REQ-007", "Leadership metrics tracked by department", "A.5.4.3", "10", "", "", "", "10", "", ""),
        ("REQ-008", "Security culture survey conducted annually", "A.5.4.4", "10", "", "", "", "10", "", ""),
        ("REQ-009", "Culture survey results actioned", "A.5.4.4", "5", "", "", "", "5", "", ""),
        ("REQ-010", "Management commitment evidence available for audit", "A.5.4", "5", "", "", "", "5", "", ""),
    ]

    row = 2
    for req in requirements:
        for col, value in enumerate(req, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if col in [5, 6, 7, 9, 10]:
                cell.fill = INPUT_FILL
        row += 1

    # Add empty rows
    for row in range(row, row + 5):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validations
    evidence_dv = DataValidation(type="list", formula1='"Yes,Partial,No"')
    ws.add_data_validation(evidence_dv)
    evidence_dv.add('E2:E20')

    impl_dv = DataValidation(type="list", formula1='"Fully Implemented,Partially Implemented,Not Implemented,Not Applicable"')
    ws.add_data_validation(impl_dv)
    impl_dv.add('F2:F20')

    remed_dv = DataValidation(type="list", formula1='"Not Required,In Progress,Planned,Completed"')
    ws.add_data_validation(remed_dv)
    remed_dv.add('J2:J20')

    # Column widths
    widths = [12, 50, 10, 8, 18, 22, 8, 10, 35, 18]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_gap_analysis_sheet(ws):
    """Create the Gap Analysis sheet."""
    ws.title = "Gap_Analysis"

    headers = [
        "Gap_ID", "Requirement_Reference", "Gap_Description", "Risk_Level",
        "Root_Cause", "Remediation_Action", "Owner", "Target_Date",
        "Status", "Progress", "Evidence_of_Closure", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    risk_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(risk_dv)
    risk_dv.add('D2:D30')

    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Closed,Deferred"')
    ws.add_data_validation(status_dv)
    status_dv.add('I2:I30')

    progress_dv = DataValidation(type="list", formula1='"0%,25%,50%,75%,100%"')
    ws.add_data_validation(progress_dv)
    progress_dv.add('J2:J30')

    # Format input rows
    for row in range(2, 31):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [10, 18, 40, 10, 30, 40, 18, 12, 12, 10, 25, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_audit_readiness_sheet(ws):
    """Create the Audit Readiness sheet."""
    ws.title = "Audit_Readiness"

    headers = [
        "Check_ID", "Audit_Requirement", "Evidence_Required", "Evidence_Location",
        "Evidence_Available", "Last_Reviewed", "Reviewer", "Status", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate audit checks
    checks = [
        ("AUD-001", "Management commitment policy exists", "ISMS-POL-A.5.4 document",
         "", "", "", "", "", ""),
        ("AUD-002", "Manager assessment records available", "A.5.4.1 completed workbooks",
         "", "", "", "", "", ""),
        ("AUD-003", "Commitment scores calculated and tracked", "A.5.4.1 Summary Scores sheet",
         "", "", "", "", "", ""),
        ("AUD-004", "Training oversight documented", "A.5.4.2 Training Oversight sheet",
         "", "", "", "", "", ""),
        ("AUD-005", "Policy violation handling documented", "A.5.4.2 Policy Violations sheet",
         "", "", "", "", "", ""),
        ("AUD-006", "Access review participation tracked", "A.5.4.2 Access Reviews sheet",
         "", "", "", "", "", ""),
        ("AUD-007", "Escalation process evidence", "A.5.4.2 Escalation Triggers sheet",
         "", "", "", "", "", ""),
        ("AUD-008", "Leadership dashboard maintained", "A.5.4.3 workbook",
         "", "", "", "", "", ""),
        ("AUD-009", "Security culture survey conducted", "A.5.4.4 Response Data sheet",
         "", "", "", "", "", ""),
        ("AUD-010", "Culture improvement actions tracked", "A.5.4.4 Action Plan sheet",
         "", "", "", "", "", ""),
    ]

    row = 2
    for check in checks:
        for col, value in enumerate(check, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if col >= 4:
                cell.fill = INPUT_FILL
        row += 1

    # Add empty rows
    for row in range(row, row + 5):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validations
    avail_dv = DataValidation(type="list", formula1='"Yes,Partial,No"')
    ws.add_data_validation(avail_dv)
    avail_dv.add('E2:E20')

    status_dv = DataValidation(type="list", formula1='"Ready,Action Required,Not Applicable"')
    ws.add_data_validation(status_dv)
    status_dv.add('H2:H20')

    # Column widths
    widths = [10, 38, 35, 30, 18, 15, 18, 18, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_trend_analysis_sheet(ws):
    """Create the Trend Analysis sheet."""
    ws.title = "Trend_Analysis"

    headers = [
        "Period", "Commitment_Score", "Training_Compliance", "Violation_Response",
        "Access_Review_Rate", "Culture_Score", "Overall_Score", "Key_Changes", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate periods
    periods = ["Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025",
               "Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"]

    row = 2
    for period in periods:
        ws.cell(row=row, column=1, value=period).border = THIN_BORDER
        for col in range(2, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        row += 1

    # Column widths
    widths = [12, 18, 20, 20, 18, 15, 15, 35, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_evidence_register_sheet(ws):
    """Create the Evidence Register sheet."""
    ws.title = "Evidence_Register"

    headers = [
        "Evidence_ID", "Evidence_Type", "Description", "Related_Requirement",
        "Date_Created", "Created_By", "Storage_Location",
        "Retention_Period", "Review_Date", "Status", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(
        type="list",
        formula1='"Assessment Record,Training Record,Violation Record,Review Record,Survey Results,Approval Record,Other"'
    )
    ws.add_data_validation(type_dv)
    type_dv.add('B2:B50')

    status_dv = DataValidation(type="list", formula1='"Current,Archived,Pending Review"')
    ws.add_data_validation(status_dv)
    status_dv.add('J2:J50')

    # Format input rows
    for row in range(2, 31):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 20, 40, 18, 15, 20, 40, 15, 15, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_approval_signoff_sheet(ws):
    """Create the Approval Sign-Off sheet."""
    ws.title = "Approval_SignOff"

    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="A.5.4 Compliance Dashboard - Approval & Sign-Off")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center")

    # Assessment details
    ws.cell(row=3, column=1, value="Assessment Period:").font = Font(bold=True)
    ws.cell(row=3, column=2).fill = INPUT_FILL
    ws.cell(row=4, column=1, value="Assessment Date:").font = Font(bold=True)
    ws.cell(row=4, column=2).fill = INPUT_FILL
    ws.cell(row=5, column=1, value="Prepared By:").font = Font(bold=True)
    ws.cell(row=5, column=2).fill = INPUT_FILL

    # Sign-off table
    ws.cell(row=7, column=1, value="APPROVAL WORKFLOW").font = SECTION_FONT

    headers = ["Role", "Name", "Date", "Signature", "Comments"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    roles = [
        "Prepared By (Security Analyst)",
        "Reviewed By (Security Manager)",
        "Validated By (HR Director)",
        "Approved By (CISO)",
    ]

    row = 9
    for role in roles:
        ws.cell(row=row, column=1, value=role).border = THIN_BORDER
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        row += 1

    # Declaration
    ws.cell(row=row + 1, column=1, value="DECLARATION").font = SECTION_FONT
    declaration = ("I confirm that this compliance dashboard accurately represents the current "
                  "status of ISO 27001:2022 Control A.5.4 (Management Responsibilities) based on "
                  "the source assessment workbooks. All gaps and remediation actions have been "
                  "identified and assigned to responsible owners.")
    ws.merge_cells(f'A{row+2}:E{row+3}')
    ws.cell(row=row + 2, column=1, value=declaration).alignment = Alignment(wrap_text=True)

    # Column widths
    widths = [35, 25, 15, 20, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_workbook():
    """Generate the complete compliance dashboard workbook."""
    logger.info("=" * 70)
    logger.info(f"Generating: {DOCUMENT_ID} - {WORKBOOK_NAME}")
    logger.info(f"Control: {CONTROL_REF}")
    logger.info("=" * 70)

    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create all sheets
    logger.info("Creating Instructions sheet...")
    create_instructions_sheet(wb.create_sheet())

    logger.info("Creating Executive_Summary sheet...")
    create_executive_summary_sheet(wb.create_sheet())

    logger.info("Creating Compliance_KPIs sheet...")
    create_compliance_kpis_sheet(wb.create_sheet())

    logger.info("Creating Compliance_Scorecard sheet...")
    create_compliance_scorecard_sheet(wb.create_sheet())

    logger.info("Creating Gap_Analysis sheet...")
    create_gap_analysis_sheet(wb.create_sheet())

    logger.info("Creating Audit_Readiness sheet...")
    create_audit_readiness_sheet(wb.create_sheet())

    logger.info("Creating Trend_Analysis sheet...")
    create_trend_analysis_sheet(wb.create_sheet())

    logger.info("Creating Evidence_Register sheet...")
    create_evidence_register_sheet(wb.create_sheet())

    logger.info("Creating Approval_SignOff sheet...")
    create_approval_signoff_sheet(wb.create_sheet())

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info("=" * 70)
    logger.info(f"SUCCESS: Workbook saved as {OUTPUT_FILENAME}")
    logger.info(f"Sheets created: {len(wb.sheetnames)}")
    logger.info("=" * 70)

    return OUTPUT_FILENAME


# =============================================================================
# ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    try:
        generate_workbook()
    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        sys.exit(1)

# =============================================================================
# QA_VERIFIED: 2026-02-04
# QA_STATUS: PASSED - NEW COMPLIANCE DASHBOARD
# QA_TOOL: Claude Code
# CHANGES: New generator created to provide proper compliance dashboard for A.5.4
#          consolidating data from .1 (commitment), .2 (oversight), .3 (leadership),
#          .4 (culture survey) into unified compliance view
# =============================================================================
