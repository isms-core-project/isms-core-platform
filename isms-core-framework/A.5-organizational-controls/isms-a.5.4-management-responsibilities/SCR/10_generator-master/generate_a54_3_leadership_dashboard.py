#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.4.3 - Leadership Dashboard
================================================================================

ISO/IEC 27001:2022 Control A.5.4: Management Responsibilities
Assessment Domain 3 of 3: Leadership Dashboard

This script generates an Excel dashboard summarizing management security
leadership metrics across the organization.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.4.3"
WORKBOOK_NAME = "Leadership Dashboard"
CONTROL_ID = "A.5.4"
CONTROL_NAME = "Management Responsibilities"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

METRIC_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions sheet."""
    ws.title = "Instructions"

    instructions = [
        ["ISMS-IMP-A.5.4.3 - Leadership Dashboard"],
        [""],
        ["PURPOSE"],
        ["This dashboard consolidates management security leadership metrics"],
        ["for executive reporting and ISMS management review."],
        [""],
        ["SHEETS"],
        ["1. Instructions - This guidance sheet"],
        ["2. Executive Summary - High-level metrics for GL"],
        ["3. By Department - Department-level breakdown"],
        ["4. Trend Analysis - Quarter-over-quarter trends"],
        ["5. Action Items - Improvement actions tracking"],
        [""],
        ["DATA SOURCES"],
        ["- ISMS-IMP-A.5.4.1 (Management Commitment Assessment)"],
        ["- ISMS-IMP-A.5.4.2 (Compliance Oversight Tracker)"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SHEETS", "DATA SOURCES"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 60


def create_executive_summary_sheet(ws):
    """Create the Executive Summary sheet."""
    ws.title = "Executive Summary"

    # Title
    ws.merge_cells('A1:E1')
    cell = ws.cell(row=1, column=1, value="Management Security Leadership - Executive Summary")
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value=f"Report Date: {GENERATED_DATE}")
    ws.cell(row=2, column=1).font = Font(italic=True)

    # Key Metrics Section
    metrics = [
        ("Overall Commitment Score", "[Score]%", "70%", "85%"),
        ("Managers Assessed", "[Count]", "-", "-"),
        ("Managers Meeting Target", "[Count]", "-", "-"),
        ("Training Oversight Compliance", "[Score]%", "90%", "95%"),
        ("Access Reviews Completed", "[Score]%", "100%", "100%"),
        ("Avg Policy Violation Response", "[Days]", "5 days", "3 days"),
    ]

    row = 4
    ws.cell(row=row, column=1, value="Key Metric").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Current").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Minimum").font = Font(bold=True)
    ws.cell(row=row, column=4, value="Target").font = Font(bold=True)

    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).font = HEADER_FONT
        ws.cell(row=row, column=col).border = THIN_BORDER

    row = 5
    for metric, current, minimum, target in metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=2, value=current).border = THIN_BORDER
        ws.cell(row=row, column=2).fill = INPUT_FILL
        ws.cell(row=row, column=3, value=minimum).border = THIN_BORDER
        ws.cell(row=row, column=4, value=target).border = THIN_BORDER
        row += 1

    # Status Summary
    row += 2
    ws.cell(row=row, column=1, value="Overall Status:").font = Font(bold=True)
    ws.cell(row=row, column=2, value="[MEETS/BELOW/EXCEEDS]").fill = INPUT_FILL

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12


def create_by_department_sheet(ws):
    """Create the By Department sheet."""
    ws.title = "By Department"

    headers = [
        "Department", "Manager_Count", "Avg_Commitment_Score",
        "Training_Compliance", "Violations_Handled", "Access_Reviews",
        "Status", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Sample departments
    departments = [
        "Executive", "Finance", "IT", "Operations", "HR",
        "Legal", "Sales", "Marketing", "Engineering", "Support"
    ]

    row = 2
    for dept in departments:
        ws.cell(row=row, column=1, value=dept).border = THIN_BORDER
        for col in range(2, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        row += 1

    # Status validation
    status_dv = DataValidation(type="list", formula1='"Exceeds,Meets,Below,N/A"')
    ws.add_data_validation(status_dv)
    status_dv.add('G2:G20')

    # Column widths
    widths = [15, 15, 20, 18, 18, 15, 10, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_trend_analysis_sheet(ws):
    """Create the Trend Analysis sheet."""
    ws.title = "Trend Analysis"

    headers = [
        "Metric", "Q1", "Q2", "Q3", "Q4", "YoY_Change", "Trend"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Metrics to track
    metrics = [
        "Overall Commitment Score (%)",
        "Managers at Target (%)",
        "Training Oversight (%)",
        "Access Review Completion (%)",
        "Violation Response Time (days)",
        "Security Culture Survey Score",
    ]

    row = 2
    for metric in metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        for col in range(2, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        row += 1

    # Trend validation
    trend_dv = DataValidation(type="list", formula1='"↑ Improving,→ Stable,↓ Declining"')
    ws.add_data_validation(trend_dv)
    trend_dv.add('G2:G20')

    # Column widths
    widths = [35, 12, 12, 12, 12, 12, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_action_items_sheet(ws):
    """Create the Action Items sheet."""
    ws.title = "Action Items"

    headers = [
        "Action_ID", "Finding", "Action_Required", "Owner",
        "Due_Date", "Status", "Completion_Date", "Evidence"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Status validation
    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Completed,Overdue,Cancelled"')
    ws.add_data_validation(status_dv)
    status_dv.add('F2:F50')

    # Format input rows
    for row in range(2, 51):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 35, 40, 20, 12, 12, 15, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_workbook():
    """Generate the complete dashboard workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()
    default_sheet = wb.active

    create_instructions_sheet(wb.create_sheet())
    create_executive_summary_sheet(wb.create_sheet())
    create_by_department_sheet(wb.create_sheet())
    create_trend_analysis_sheet(wb.create_sheet())
    create_action_items_sheet(wb.create_sheet())

    wb.remove(default_sheet)

    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - INITIAL CREATION
# QA_TOOL: Claude Code
# CHANGES: Initial generator creation for A.5.4 Management Responsibilities
# =============================================================================
