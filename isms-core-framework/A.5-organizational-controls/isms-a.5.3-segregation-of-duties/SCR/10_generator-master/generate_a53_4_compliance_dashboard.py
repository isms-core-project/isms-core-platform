#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.3.4 - Compliance Dashboard
================================================================================

ISO/IEC 27001:2022 Control A.5.3: Segregation of Duties
Assessment Workbook 4 of 4: Compliance Dashboard

This script generates an Excel workbook providing executive-level dashboard
for SoD compliance monitoring, KPI tracking, and audit evidence compilation.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.3.4"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.5.3"
CONTROL_NAME = "Segregation of Duties"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

METRIC_FONT = Font(bold=True, size=16)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Traffic light colors
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# DATA CONSTANTS
# =============================================================================
KPI_NAMES = [
    "Conflict Identification Rate",
    "Resolution Rate",
    "Mean Time to Resolution (days)",
    "Exception Ratio",
    "Control Effectiveness",
    "Validation Coverage"
]

REMEDIATION_STATUSES = ["Not Started", "In Progress", "Completed", "Cancelled", "Overdue"]

EXCEPTION_CONTROL_STATUS = ["Yes", "No", "Partial"]

DEPARTMENTS = [
    "Executive",
    "Finance",
    "IT",
    "Operations",
    "HR",
    "Legal",
    "Sales",
    "Marketing",
    "Engineering",
    "Support",
    "Procurement",
    "Security"
]

EVIDENCE_STATUSES = ["Ready", "In Progress", "Missing"]

DATA_SOURCES = [
    ("ISMS-IMP-A.5.3.1", "SoD Matrix Assessment", "Monthly"),
    ("ISMS-IMP-A.5.3.2", "Conflict Analysis", "Quarterly"),
    ("ISMS-IMP-A.5.3.3", "Role-Function Mapping", "Quarterly"),
    ("IAM System", "Access Rights Export", "Monthly"),
    ("HR System", "Organisation Chart", "Monthly"),
]

REFRESH_FREQUENCIES = ["Daily", "Weekly", "Monthly", "Quarterly"]


# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def apply_header_style(cell):
    """Apply standard header styling to a cell."""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGNMENT
    cell.border = THIN_BORDER


def apply_input_style(cell):
    """Apply input cell styling."""
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER


def create_executive_dashboard_sheet(ws):
    """Create the Executive Dashboard sheet."""
    ws.title = "Executive_Dashboard"

    # Title
    cell = ws.cell(row=1, column=1, value="Segregation of Duties - Executive Dashboard")
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    ws.merge_cells('A1:F1')

    # Date
    ws.cell(row=2, column=1, value=f"As of: {GENERATED_DATE}").font = Font(italic=True)

    # Traffic Light Summary Section
    ws.cell(row=4, column=1, value="COMPLIANCE STATUS").font = Font(bold=True, size=12)

    metrics = [
        ("Overall Compliance Score", ">95%", ""),
        ("Critical Conflicts Open", "0", ""),
        ("Remediation On Track", ">90%", ""),
        ("Active Exceptions", "<5", ""),
        ("Days Since Assessment", "<90", ""),
    ]

    headers = ["Metric", "Target", "Current", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_header_style(cell)

    for row_idx, (metric, target, current) in enumerate(metrics, 6):
        ws.cell(row=row_idx, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=target).border = THIN_BORDER
        cell_current = ws.cell(row=row_idx, column=3, value=current)
        cell_current.fill = INPUT_FILL
        cell_current.border = THIN_BORDER
        cell_status = ws.cell(row=row_idx, column=4, value="")
        cell_status.fill = INPUT_FILL
        cell_status.border = THIN_BORDER

    # Key Statistics Section
    ws.cell(row=12, column=1, value="KEY STATISTICS").font = Font(bold=True, size=12)

    stats = [
        ("Total Roles Assessed", ""),
        ("Total Conflicts Identified", ""),
        ("Conflicts Resolved", ""),
        ("Conflicts In Remediation", ""),
        ("Active Exceptions", ""),
    ]

    for row_idx, (stat, value) in enumerate(stats, 13):
        ws.cell(row=row_idx, column=1, value=stat).border = THIN_BORDER
        cell_val = ws.cell(row=row_idx, column=2, value=value)
        cell_val.fill = INPUT_FILL
        cell_val.border = THIN_BORDER
        cell_val.font = METRIC_FONT

    # Executive Summary Section
    ws.cell(row=19, column=1, value="EXECUTIVE SUMMARY").font = Font(bold=True, size=12)
    ws.merge_cells('A20:F25')
    cell = ws.cell(row=20, column=1, value="[Enter executive summary narrative here]")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.fill = INPUT_FILL

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_kpi_scorecard_sheet(ws):
    """Create the KPI Scorecard sheet."""
    ws.title = "KPI_Scorecard"

    headers = [
        "KPI_Name", "Target", "Q1", "Q2", "Q3", "Q4", "YTD", "Status", "Trend"
    ]

    widths = [30, 12, 10, 10, 10, 10, 10, 15, 10]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

    # Pre-populate KPIs
    kpi_targets = [
        ("Conflict Identification Rate", "Baseline"),
        ("Resolution Rate (%)", "80"),
        ("Mean Time to Resolution (days)", "90"),
        ("Exception Ratio (%)", "20"),
        ("Control Effectiveness (%)", "70"),
        ("Validation Coverage (%)", "100"),
    ]

    for row_idx, (kpi, target) in enumerate(kpi_targets, 2):
        ws.cell(row=row_idx, column=1, value=kpi).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=target).border = THIN_BORDER

        # Input cells for quarters
        for col in range(3, 7):
            apply_input_style(ws.cell(row=row_idx, column=col))

        # YTD formula (average of quarters)
        cell = ws.cell(row=row_idx, column=7,
                       value=f"=IF(COUNT(C{row_idx}:F{row_idx})>0,AVERAGE(C{row_idx}:F{row_idx}),\"\")")
        cell.fill = FORMULA_FILL
        cell.border = THIN_BORDER

        # Status and Trend (input)
        apply_input_style(ws.cell(row=row_idx, column=8))
        apply_input_style(ws.cell(row=row_idx, column=9))


def create_conflict_status_sheet(ws):
    """Create the Conflict Status sheet."""
    ws.title = "Conflict_Status"

    # Section 1: By Conflict Type
    ws.cell(row=1, column=1, value="CONFLICTS BY TYPE").font = Font(bold=True, size=12)

    type_headers = ["Conflict_Type", "Total", "Open", "Mitigated", "Resolved", "Accepted", "% Resolved"]
    for col, header in enumerate(type_headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        apply_header_style(cell)

    types = ["X (Hard)", "C (Conditional)", "M (Monitor)", "Total"]
    for row_idx, ctype in enumerate(types, 3):
        ws.cell(row=row_idx, column=1, value=ctype).border = THIN_BORDER
        for col in range(2, 7):
            apply_input_style(ws.cell(row=row_idx, column=col))
        # % Resolved formula
        cell = ws.cell(row=row_idx, column=7,
                       value=f"=IF(B{row_idx}>0,(E{row_idx}/(B{row_idx}-F{row_idx}))*100,0)")
        cell.fill = FORMULA_FILL
        cell.border = THIN_BORDER

    # Section 2: By Process Domain
    ws.cell(row=9, column=1, value="CONFLICTS BY PROCESS").font = Font(bold=True, size=12)

    process_headers = ["Process_Domain", "Total", "Critical", "High", "Medium", "Low"]
    for col, header in enumerate(process_headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        apply_header_style(cell)

    processes = ["Financial", "IT Operations", "HR", "Procurement", "Security", "Change Management"]
    for row_idx, process in enumerate(processes, 11):
        ws.cell(row=row_idx, column=1, value=process).border = THIN_BORDER
        for col in range(2, 7):
            apply_input_style(ws.cell(row=row_idx, column=col))

    # Section 3: Aging Analysis
    ws.cell(row=19, column=1, value="CONFLICT AGING").font = Font(bold=True, size=12)

    age_headers = ["Age_Bucket", "Count", "% of Total"]
    for col, header in enumerate(age_headers, 1):
        cell = ws.cell(row=20, column=col, value=header)
        apply_header_style(cell)

    buckets = ["<30 days", "30-60 days", "60-90 days", ">90 days"]
    for row_idx, bucket in enumerate(buckets, 21):
        ws.cell(row=row_idx, column=1, value=bucket).border = THIN_BORDER
        apply_input_style(ws.cell(row=row_idx, column=2))
        apply_input_style(ws.cell(row=row_idx, column=3))

    # Column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15


def create_remediation_progress_sheet(ws):
    """Create the Remediation Progress sheet."""
    ws.title = "Remediation_Progress"

    headers = [
        "Remediation_ID", "Gap_ID", "Owner", "Target_Date",
        "Status", "Days_Remaining", "Escalation_Status"
    ]

    widths = [15, 15, 25, 15, 15, 15, 20]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

    # Data validation for status
    status_dv = DataValidation(type="list", formula1=f'"{",".join(REMEDIATION_STATUSES)}"')
    ws.add_data_validation(status_dv)
    status_dv.add('E2:E100')

    # Format input rows with Days_Remaining formula
    for row in range(2, 51):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            if col == 6:  # Days_Remaining formula
                cell.value = f'=IF(D{row}="","",D{row}-TODAY())'
                cell.fill = FORMULA_FILL
            else:
                apply_input_style(cell)
            cell.border = THIN_BORDER


def create_exception_monitoring_sheet(ws):
    """Create the Exception Monitoring sheet."""
    ws.title = "Exception_Monitoring"

    headers = [
        "Exception_ID", "Gap_ID", "Justification", "Compensating_Controls",
        "Expiry_Date", "Days_Until_Expiry", "Last_Review", "Control_Effective", "Status"
    ]

    widths = [15, 15, 35, 40, 15, 18, 15, 15, 12]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

    # Data validations
    eff_dv = DataValidation(type="list", formula1=f'"{",".join(EXCEPTION_CONTROL_STATUS)}"')
    ws.add_data_validation(eff_dv)
    eff_dv.add('H2:H100')

    status_dv = DataValidation(type="list", formula1='"Active,Expired,Revoked"')
    ws.add_data_validation(status_dv)
    status_dv.add('I2:I100')

    # Format input rows with Days_Until_Expiry formula
    for row in range(2, 51):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            if col == 6:  # Days_Until_Expiry formula
                cell.value = f'=IF(E{row}="","",E{row}-TODAY())'
                cell.fill = FORMULA_FILL
            else:
                apply_input_style(cell)
            cell.border = THIN_BORDER


def create_trend_analysis_sheet(ws):
    """Create the Trend Analysis sheet."""
    ws.title = "Trend_Analysis"

    headers = [
        "Period", "Total_Conflicts", "Critical_Conflicts", "Resolved_Count",
        "MTTR_Days", "Compliance_Pct", "Exceptions_Active", "Trend_vs_Prior"
    ]

    widths = [12, 15, 18, 15, 12, 15, 18, 15]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

    # Pre-populate periods
    periods = [
        "2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4",
        "2026-Q1", "2026-Q2", "2026-Q3", "2026-Q4"
    ]

    for row_idx, period in enumerate(periods, 2):
        ws.cell(row=row_idx, column=1, value=period).border = THIN_BORDER
        for col in range(2, 8):
            apply_input_style(ws.cell(row=row_idx, column=col))

        # Trend formula (compare to prior row)
        if row_idx > 2:
            cell = ws.cell(row=row_idx, column=8,
                           value=f'=IF(AND(F{row_idx}<>"",F{row_idx-1}<>""),F{row_idx}-F{row_idx-1},"")')
            cell.fill = FORMULA_FILL
        else:
            apply_input_style(ws.cell(row=row_idx, column=8))
        ws.cell(row=row_idx, column=8).border = THIN_BORDER


def create_department_view_sheet(ws):
    """Create the Department View sheet."""
    ws.title = "Department_View"

    headers = [
        "Department", "Total_Roles", "Conflicts_Identified", "Conflicts_Resolved",
        "Active_Exceptions", "Compliance_Pct", "Status"
    ]

    widths = [20, 12, 20, 18, 18, 15, 12]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

    # Pre-populate departments
    for row_idx, dept in enumerate(DEPARTMENTS, 2):
        ws.cell(row=row_idx, column=1, value=dept).border = THIN_BORDER
        for col in range(2, 6):
            apply_input_style(ws.cell(row=row_idx, column=col))

        # Compliance_Pct formula
        cell = ws.cell(row=row_idx, column=6,
                       value=f'=IF(B{row_idx}=0,"N/A",(B{row_idx}-C{row_idx}+D{row_idx})/B{row_idx}*100)')
        cell.fill = FORMULA_FILL
        cell.border = THIN_BORDER

        # Status formula
        cell = ws.cell(row=row_idx, column=7,
                       value=f'=IF(F{row_idx}="N/A","N/A",IF(F{row_idx}>=95,"GREEN",IF(F{row_idx}>=85,"YELLOW","RED")))')
        cell.fill = FORMULA_FILL
        cell.border = THIN_BORDER


def create_audit_evidence_sheet(ws):
    """Create the Audit Evidence sheet."""
    ws.title = "Audit_Evidence"

    headers = [
        "Evidence_Item", "Document_ID", "Location", "Date", "Status", "Notes"
    ]

    widths = [35, 20, 50, 15, 15, 40]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

    # Data validation for status
    status_dv = DataValidation(type="list", formula1=f'"{",".join(EVIDENCE_STATUSES)}"')
    ws.add_data_validation(status_dv)
    status_dv.add('E2:E50')

    # Pre-populate evidence items
    evidence_items = [
        ("SoD Matrix (current)", "ISMS-IMP-A.5.3.1", "[SharePoint path]", "", "", ""),
        ("Gap Analysis", "ISMS-IMP-A.5.3.1", "[SharePoint path]", "", "", ""),
        ("Conflict Analysis", "ISMS-IMP-A.5.3.2", "[SharePoint path]", "", "", ""),
        ("Role-Function Mapping", "ISMS-IMP-A.5.3.3", "[SharePoint path]", "", "", ""),
        ("Exception Approvals", "Evidence folder", "[SharePoint path]", "", "", ""),
        ("Compensating Control Evidence", "Evidence folder", "[SharePoint path]", "", "", ""),
        ("Historical Dashboards", "Archive", "[SharePoint path]", "", "", ""),
        ("Management Review Minutes", "ISMS Records", "[SharePoint path]", "", "", ""),
    ]

    for row_idx, evidence in enumerate(evidence_items, 2):
        for col_idx, value in enumerate(evidence, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx >= 3:
                apply_input_style(cell)
            cell.border = THIN_BORDER


def create_data_sources_sheet(ws):
    """Create the Data Sources sheet."""
    ws.title = "Data_Sources"

    headers = [
        "Source_Name", "Document_ID", "File_Location", "Last_Refresh",
        "Refresh_Frequency", "Responsible_Party"
    ]

    widths = [25, 20, 50, 15, 15, 25]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

    # Data validation for frequency
    freq_dv = DataValidation(type="list", formula1=f'"{",".join(REFRESH_FREQUENCIES)}"')
    ws.add_data_validation(freq_dv)
    freq_dv.add('E2:E20')

    # Pre-populate data sources
    for row_idx, (name, doc_id, freq) in enumerate(DATA_SOURCES, 2):
        ws.cell(row=row_idx, column=1, value=name).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=doc_id).border = THIN_BORDER
        apply_input_style(ws.cell(row=row_idx, column=3))
        apply_input_style(ws.cell(row=row_idx, column=4))
        ws.cell(row=row_idx, column=5, value=freq).border = THIN_BORDER
        apply_input_style(ws.cell(row=row_idx, column=6))


def generate_workbook():
    """Generate the complete dashboard workbook."""
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Create all sheets
    create_executive_dashboard_sheet(wb.create_sheet())
    create_kpi_scorecard_sheet(wb.create_sheet())
    create_conflict_status_sheet(wb.create_sheet())
    create_remediation_progress_sheet(wb.create_sheet())
    create_exception_monitoring_sheet(wb.create_sheet())
    create_trend_analysis_sheet(wb.create_sheet())
    create_department_view_sheet(wb.create_sheet())
    create_audit_evidence_sheet(wb.create_sheet())
    create_data_sources_sheet(wb.create_sheet())

    # Remove default sheet
    wb.remove(default_sheet)

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")

    return OUTPUT_FILENAME


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - INITIAL CREATION
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.3 Segregation of Duties control
# =============================================================================
