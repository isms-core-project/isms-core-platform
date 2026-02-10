#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.34.5 - Data Protection Impact Assessment (DPIA) Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.34: Privacy and Protection of PII
Assessment Domain 5 of 7: Data Protection Impact Assessment (DPIA) - GDPR Art. 35

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific data processing operations, risk assessment methodologies,
and GDPR Article 35 DPIA requirements.

Key customization areas:
1. DPIA trigger criteria (match your regulatory obligations and risk profile)
2. High-risk processing scenarios (adapt to your industry and data types)
3. Risk assessment methodology (align with organizational risk framework)
4. Mitigation measures catalog (specific to your technical capabilities)
5. Stakeholder consultation processes (based on organizational structure)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.34 Privacy and Protection of PII Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for conducting
Data Protection Impact Assessments (DPIAs) per GDPR Article 35, ISO/IEC 29134:2017
guidelines, and Article 29 Working Party WP248 recommendations.

**Purpose:**
Enables systematic identification of high-risk data processing, structured DPIA
execution with multi-criteria risk assessment, and evidence-based validation of
risk mitigation measures to demonstrate ISO 27001:2022 Control A.5.34 compliance
and meet GDPR Article 35 mandatory DPIA obligations.

**Assessment Scope:**
- DPIA trigger assessment (9 GDPR Article 35 criteria + WP248 guidelines)
- DPIA register maintenance (catalog of all completed and ongoing DPIAs)
- Systematic risk identification (data subject rights and freedoms impact)
- Multi-criteria risk assessment (likelihood × severity × data subject impact)
- Mitigation measure effectiveness evaluation (residual risk calculation)
- DPO consultation documentation (GDPR Article 35(2) requirement)
- Data subject or representative consultation (GDPR Article 35(9) where appropriate)
- Supervisory authority pre-consultation for high residual risks (GDPR Article 36)
- Gap analysis and remediation tracking
- Evidence collection for audit readiness
- Compliance dashboard with DPIA coverage metrics

**Generated Workbook Structure:**
1. Instructions & Legend - DPIA execution guidance and GDPR Article 35 requirements
2. Trigger Assessment - 9-criteria evaluation to determine DPIA necessity
3. DPIA Register - Comprehensive catalog of all DPIAs with status tracking
4. Processing Description - Detailed documentation of high-risk processing activity
5. Risk Assessment - Systematic identification and evaluation of risks to data subjects
6. Mitigation Measures - Control implementation tracking with residual risk calculation
7. Stakeholder Consultation - DPO, data subject, and supervisory authority consultation log
8. Gap Analysis - Unmitigated or partially mitigated risks requiring action
9. Evidence Repository - Audit evidence tracking and documentation linkage
10. Dashboard - DPIA coverage, risk trends, and compliance metrics

**Key Features:**
- Data validation with GDPR Article 35 trigger criteria dropdowns
- Conditional formatting for risk levels (Critical/High/Medium/Low)
- Automated DPIA necessity determination based on 9 trigger criteria
- Risk matrix with likelihood × severity × impact calculation
- Residual risk auto-calculation post-mitigation
- Protected formulas with unprotected input cells
- DPO consultation requirement flagging (mandatory per GDPR Article 35(2))
- Supervisory authority pre-consultation alerts for high residual risks
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Dashboard with DPIA completion rates and risk distribution

**Integration:**
This assessment is Domain 5 of the A.5.34 Privacy assessment suite (7 domains).
Results feed into ISMS-IMP-A.5.34.7 Privacy Compliance Dashboard for
consolidated privacy program oversight and regulatory reporting. DPIAs integrate
with A.5.34.1 (PII Inventory), A.5.34.2 (Legal Basis), and A.5.34.4 (TOMs) for
holistic risk assessment.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a5345_dpia_assessment.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a5345_dpia_assessment.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a5345_dpia_assessment.py --date 20250128

Output:
    File: ISMS_A_5_34_5_DPIA_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize DPIA trigger criteria to match applicable regulations
    2. Identify all high-risk processing activities requiring DPIA (use Sheet 2 trigger assessment)
    3. Create DPIA register entries for mandatory DPIAs (new systems, profiling, automated decisions)
    4. Document detailed processing descriptions including data flows and retention
    5. Conduct systematic risk assessment for each DPIA (risks to data subject rights/freedoms)
    6. Identify and document mitigation measures (technical and organizational)
    7. Calculate residual risk post-mitigation (must be acceptable or trigger Art. 36 consultation)
    8. Consult DPO on DPIA execution (MANDATORY per GDPR Article 35(2))
    9. Seek data subject or representative views where appropriate (GDPR Article 35(9))
    10. Initiate supervisory authority pre-consultation if high residual risk remains (GDPR Article 36)
    11. Review DPIA annually or when processing changes significantly
    12. Collect and link audit evidence (risk analysis, DPO advice, SA consultation)
    13. Review dashboard metrics with Privacy Committee
    14. Obtain stakeholder approvals
    15. Feed results into A.5.34.7 Privacy Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.34
Assessment Domain:    5 of 7 (Data Protection Impact Assessment - DPIA)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.34: Privacy and Protection of PII Policy (Governance)
    - ISMS-CTX-A.5.34: Privacy Regulatory Landscape Reference (Context)
    - ISMS-IMP-A.5.34.1: PII Identification Assessment (Domain 1)
    - ISMS-IMP-A.5.34.2: Legal Basis Assessment (Domain 2)
    - ISMS-IMP-A.5.34.3: DSR Management Assessment (Domain 3)
    - ISMS-IMP-A.5.34.4: TOMs Assessment (Domain 4)
    - ISMS-IMP-A.5.34.5: DPIA Implementation Guide (Part 1 & 2)
    - ISMS-IMP-A.5.34.6: Cross-Border Transfer Assessment (Domain 6)
    - ISMS-IMP-A.5.34.7: Privacy Compliance Dashboard (Consolidation)
    - ISO/IEC 29134:2017: DPIA Guidelines (External Reference)
    - Article 29 WP248: DPIA Guidelines (External Reference)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 2025-01-28
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.34.5 specification
    - Supports comprehensive DPIA execution per GDPR Article 35
    - Integrated dashboard for DPIA coverage and risk monitoring
    - Prepares for consolidation into A.5.34.7 Privacy Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**GDPR Article 35 Mandatory Requirements:**
GDPR Article 35(1) mandates DPIAs when processing is "likely to result in high risk
to the rights and freedoms of natural persons." Article 35(3) lists mandatory scenarios:
(a) systematic and extensive profiling with legal/significant effects, (b) large-scale
processing of special category data (Art. 9) or criminal offense data (Art. 10),
(c) systematic monitoring of publicly accessible areas on large scale. Supervisory
authorities publish DPIA lists (Art. 35(4) mandatory, Art. 35(5) optional).

**Article 29 Working Party WP248 Criteria:**
WP248 guidelines identify 9 DPIA trigger criteria: evaluation/scoring, automated
decision-making, systematic monitoring, sensitive data, large-scale processing,
data matching/combining, vulnerable data subjects, innovative technology, data
transfer outside EU or blocking data subject rights. Meeting 2+ criteria triggers DPIA.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 and GDPR requirements.
Auditors will expect documented DPIA necessity assessments, complete risk analyses,
evidence of DPO consultation (MANDATORY), proof of data subject consultation where
appropriate, supervisory authority pre-consultation records for high residual risks,
and regular DPIA reviews when processing changes.

**Data Protection:**
Assessment workbooks contain sensitive information including:
- Complete inventory of high-risk processing activities
- Detailed risk analyses revealing organizational vulnerabilities
- Mitigation plans with implementation timelines (security-sensitive)
- DPO advice and supervisory authority consultation records (privileged communications)

Handle in accordance with your organization's data classification policies.
Restrict access to DPO, Legal, CISO, Privacy Team, and authorized DPIA conductors.

**Maintenance:**
Review and update assessment:
- Annually: Review all active DPIAs per GDPR Article 35(11) requirement
- Triggered: Processing changes significantly, new technologies deployed, regulatory guidance updates
- Continuous: Monitor for new mandatory DPIA scenarios (supervisory authority lists)
- Quarterly: DPIA register updates for new projects and initiatives

**Quality Assurance:**
Have DPO validate all DPIA executions before deployment (MANDATORY per GDPR Article 35(2)).
Involve Information Security, Legal, Business Owners, and technical teams in risk
assessment. Consider independent external review for complex or controversial processing.

**Regulatory Alignment:**
This assessment supports compliance with:
- GDPR (EU General Data Protection Regulation) - Art. 35, 36 (DPIA and Prior Consultation)
- ISO/IEC 27001:2022 - Control A.5.34 (Privacy and Protection of PII)
- ISO/IEC 29134:2017 - Privacy Impact Assessment Guidelines
- Article 29 Working Party WP248 - DPIA Guidelines (Revised)
- ISO/IEC 27701:2019 - Privacy Information Management (if applicable)

**DPO Consultation Requirement:**
GDPR Article 35(2) MANDATES seeking DPO advice during DPIA execution. This is NOT
optional. Document DPO involvement, advice received, and how advice was addressed.
DPO must have sufficient time and information to provide meaningful input.

**Supervisory Authority Prior Consultation (Article 36):**
If DPIA indicates high residual risk despite mitigation, controller MUST consult
supervisory authority BEFORE processing (GDPR Article 36(1)). SA has 8 weeks to
respond (extensible to 14 weeks). Processing cannot commence until SA provides
written advice. Document all SA consultations thoroughly.

================================================================================
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import PieChart, BarChart, Reference
from datetime import datetime
import sys

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.34.5"
WORKBOOK_NAME = "Data Protection Impact Assessment (DPIA)"
CONTROL_ID = "A.5.34"
CONTROL_NAME = "Privacy and Protection of PII"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# Color palette (consistent with A.5.34 assessment suite)
COLORS = {
    'header_blue': 'FF003366',
    'header_orange': 'FFEB9C00', 
    'header_green': 'FF70AD47',
    'header_red': 'FFEB9C00',
    'header_gold': 'FFED7D31',
    'header_cyan': 'FF00B0F0',
    'white': 'FFFFFFFF',
    'black': 'FF000000',
    'light_green': 'FFC6EFCE',
    'dark_green': 'FF006100',
    'light_yellow': 'FFFFD966',
    'dark_orange': 'FF9C5700',
    'light_orange': 'FFFFA500',
    'light_red': 'FFFF6666',
    'dark_red': 'FF9C0006',
    'very_light_red': 'FFFFE6E6',
    'very_light_yellow': 'FFFFF3CD',
    'very_light_green': 'FFEBF8E9',
    'very_light_orange': 'FFFFF2CC',
    'light_blue': 'FFB4C6E7',
    'dark_blue': 'FF002060',
    'light_gray': 'FFD9D9D9',
}


def style_header_row(ws, row_num, color_hex, num_columns):
    """Apply consistent header styling to row."""
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(name='Calibri', size=11, bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color=COLORS['black']),
            right=Side(style='thin', color=COLORS['black']),
            top=Side(style='thin', color=COLORS['black']),
            bottom=Side(style='thin', color=COLORS['black'])
        )


def add_dropdown(ws, cell_range, options, error_msg, allow_blank=False):
    """Add dropdown data validation."""
    dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=allow_blank)
    dv.error = error_msg
    dv.errorStyle = 'stop'
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_risk_level_formatting(ws, column_letter, start_row, end_row):
    """Apply conditional formatting for risk levels (Low/Medium/High/Critical)."""
    cell_range = f'{column_letter}{start_row}:{column_letter}{end_row}'
    
    # Low: Green
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator='equal', formula=['"Low"'], 
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid'),
                   font=Font(color=COLORS['dark_green']))
    )
    
    # Medium: Yellow
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator='equal', formula=['"Medium"'], 
                   fill=PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid'),
                   font=Font(color=COLORS['dark_orange']))
    )
    
    # High: Orange
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator='equal', formula=['"High"'], 
                   fill=PatternFill(start_color=COLORS['light_orange'], end_color=COLORS['light_orange'], fill_type='solid'),
                   font=Font(color=COLORS['dark_red']))
    )
    
    # Critical: Red
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator='equal', formula=['"Critical"'], 
                   fill=PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid'),
                   font=Font(color=COLORS['dark_red'], bold=True))
    )


def protect_sheet(ws, password=None):
    """Protect sheet with standard settings."""
    ws.protection.sheet = True
    if password:
        ws.protection.password = password
    ws.protection.enable()
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertColumns = False
    ws.protection.insertRows = False
    ws.protection.deleteColumns = False
    ws.protection.deleteRows = False
    ws.protection.sort = True
    ws.protection.autoFilter = True
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True


def create_instructions_sheet(ws):
    """Create Sheet 1: Instructions & Legend (pre-populated reference material)."""
    
    # Title block
    ws.merge_cells('A1:P3')
    title_cell = ws['A1']
    title_cell.value = "DPIA Assessment - Instructions & Legend"
    title_cell.font = Font(name='Calibri', size=16, bold=True, color=COLORS['white'])
    title_cell.fill = PatternFill(start_color=COLORS['header_blue'], end_color=COLORS['header_blue'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    for col in range(1, 17):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Overview section
    current_row = 5
    ws[f'A{current_row}'] = "WORKBOOK OVERVIEW"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells(f'A{current_row}:P{current_row}')
    current_row += 1
    
    overview_text = [
        "This workbook implements GDPR Article 35 Data Protection Impact Assessments (DPIAs).",
        "A DPIA is required when processing is likely to result in HIGH RISK to individuals' rights and freedoms.",
        "",
        "Complete sheets in order:",
        "  1. Instructions (this sheet) - Read first",
        "  2. Trigger Assessment - Determine which processing activities need DPIA",
        "  3. DPIA Register - Track all DPIAs from initiation to completion",
        "  4. Risk Assessment - Evaluate risks to data subjects (necessity, proportionality, impact)",
        "  5. Mitigation Measures - Document controls to reduce risks",
        "  6. Stakeholder Consultation - Record DPO and data subject consultations (Art. 35(9))",
        "  7. Gap Analysis - Calculate residual risk after mitigation",
        "  8. Dashboard - View auto-calculated compliance metrics",
        "",
        "CRITICAL TIMING: GDPR Art. 35(1) requires DPIAs 'prior to processing' - complete BEFORE data processing begins!",
    ]
    
    for line in overview_text:
        ws[f'A{current_row}'] = line
        ws[f'A{current_row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{current_row}:P{current_row}')
        current_row += 1
    
    # GDPR Article 35(3) Mandatory Triggers
    current_row += 2
    ws[f'A{current_row}'] = "GDPR ARTICLE 35(3) - MANDATORY DPIA TRIGGERS"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells(f'A{current_row}:P{current_row}')
    current_row += 1
    
    triggers = [
        "(a) Systematic and extensive evaluation of personal aspects including profiling with legal/significant effects",
        "(b) Large-scale processing of special category data (Art. 9) or criminal convictions (Art. 10)",
        "(c) Systematic monitoring of publicly accessible areas on a large scale (e.g., CCTV surveillance)",
        "",
        "ADDITIONAL WP248 TRIGGERS (Article 29 Working Party Guidelines):",
        "• Innovative technologies or novel applications",
        "• Processing that prevents data subjects from exercising rights or using services",
        "• Large-scale processing (>5'000 data subjects or substantial proportion of population)",
        "• Matching/combining datasets from different sources beyond original purpose",
        "• Processing vulnerable data subjects (children, employees, asylum seekers, mentally ill)",
        "• Cross-border data transfers outside EEA without adequacy decision",
        "• Automated decision-making with potential for significant harm",
    ]
    
    for line in triggers:
        ws[f'A{current_row}'] = line
        ws[f'A{current_row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{current_row}:P{current_row}')
        current_row += 1
    
    # Risk Assessment Methodology
    current_row += 2
    ws[f'A{current_row}'] = "RISK ASSESSMENT METHODOLOGY"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells(f'A{current_row}:P{current_row}')
    current_row += 1
    
    ws[f'A{current_row}'] = "Likelihood Scale (1-5):"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    likelihood_scale = [
        "1 - Rare: <5% probability, multiple unlikely failures required",
        "2 - Unlikely: 5-25% probability, specific vulnerabilities must be exploited",
        "3 - Possible: 25-50% probability, known vulnerabilities not actively exploited",
        "4 - Likely: 50-75% probability, common attack vectors with known exploits",
        "5 - Almost Certain: >75% probability, inevitable without controls",
    ]
    for line in likelihood_scale:
        ws[f'A{current_row}'] = line
        ws.merge_cells(f'A{current_row}:P{current_row}')
        current_row += 1
    
    current_row += 1
    ws[f'A{current_row}'] = "Impact Scale (1-5):"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    impact_scale = [
        "1 - Negligible: No material harm, minor inconvenience",
        "2 - Minor: Limited, reversible harm (password reset, temporary disruption)",
        "3 - Moderate: Significant but manageable harm (financial loss <CHF 1'000, reputational damage)",
        "4 - Major: Severe harm with lasting consequences (identity theft, job loss, significant financial damage)",
        "5 - Severe: Catastrophic, potentially irreversible (physical harm, life-altering discrimination, severe trauma)",
    ]
    for line in impact_scale:
        ws[f'A{current_row}'] = line
        ws.merge_cells(f'A{current_row}:P{current_row}')
        current_row += 1
    
    current_row += 1
    ws[f'A{current_row}'] = "Risk Score = Likelihood × Impact (range: 1-25)"
    ws[f'A{current_row}'].font = Font(bold=True)
    ws.merge_cells(f'A{current_row}:P{current_row}')
    current_row += 1
    
    risk_levels = [
        "• Low (1-6): Standard controls adequate, no special approvals required",
        "• Medium (8-12): Enhanced controls required, regular monitoring",
        "• High (15-16): Comprehensive mitigation + senior management approval required",
        "• Critical (20-25): Supervisory authority consultation required (Art. 36), fundamental redesign needed",
    ]
    for line in risk_levels:
        ws[f'A{current_row}'] = line
        ws.merge_cells(f'A{current_row}:P{current_row}')
        current_row += 1
    
    # Color Coding Legend
    current_row += 2
    ws[f'A{current_row}'] = "COLOR CODING LEGEND"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells(f'A{current_row}:P{current_row}')
    current_row += 1
    
    color_legend = [
        ("Green", COLORS['light_green'], "Low risk, DPIA not required, controls adequate, compliant"),
        ("Yellow", COLORS['light_yellow'], "Medium risk, DPIA uncertain (DPO consultation), monitoring required"),
        ("Orange", COLORS['light_orange'], "High risk, enhanced controls required, management approval needed"),
        ("Red", COLORS['light_red'], "Critical risk, DPIA mandatory, supervisory authority consultation, immediate action"),
    ]
    
    for color_name, color_hex, description in color_legend:
        ws[f'A{current_row}'] = color_name
        ws[f'A{current_row}'].fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
        ws[f'B{current_row}'] = description
        ws[f'B{current_row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'B{current_row}:P{current_row}')
        current_row += 1
    
    # Evidence Requirements
    current_row += 2
    ws[f'A{current_row}'] = "EVIDENCE REQUIREMENTS"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells(f'A{current_row}:P{current_row}')
    current_row += 1
    
    evidence = [
        "• System architecture diagrams (data flow diagrams showing PII processing paths)",
        "• Data processing agreements (DPAs) with processors, joint controller agreements",
        "• Privacy notices provided to data subjects (demonstrate transparency)",
        "• Security control implementation documentation (encryption, access controls, monitoring)",
        "• DPO consultation records (emails, meeting minutes, formal opinions)",
        "• Supervisory authority correspondence (if Art. 36 consultation required)",
        "• Legal basis justifications (consent records, contract necessity, legitimate interests balancing test)",
        "• Data subject consultation evidence (surveys, focus groups, public consultations)",
    ]
    for line in evidence:
        ws[f'A{current_row}'] = line
        ws.merge_cells(f'A{current_row}:P{current_row}')
        current_row += 1
    
    # Common Pitfalls
    current_row += 2
    ws[f'A{current_row}'] = "⚠️ COMMON PITFALLS TO AVOID"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells(f'A{current_row}:P{current_row}')
    current_row += 1
    
    pitfalls = [
        "❌ Starting DPIA too late (after processing already began) → GDPR violation",
        "❌ Conducting DPIA in isolation without stakeholder input → Incomplete risk assessment",
        "❌ Generic risk descriptions ('data breach could occur') → Useless for mitigation planning",
        "❌ Failing to update DPIA when processing changes materially → Non-compliance",
        "❌ Not consulting DPO or supervisory authority when required → Art. 35(2) and Art. 36 violations",
        "❌ Skipping DPIA because 'we've always done it this way' → Not a valid exemption",
        "❌ Underestimating likelihood/impact to avoid mitigation work → Audit finding, potential fine",
        "❌ Confusing 'necessary' with 'convenient' → Proportionality failure",
    ]
    for line in pitfalls:
        ws[f'A{current_row}'] = line
        ws.merge_cells(f'A{current_row}:P{current_row}')
        current_row += 1
    
    # Lock entire sheet
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)
    
    protect_sheet(ws)


def create_trigger_assessment_sheet(ws):
    """Create Sheet 2: DPIA Trigger Assessment."""
    
    headers = [
        ("A1", "Processing Activity ID", 20),
        ("B1", "Processing Activity Name", 30),
        ("C1", "System/Application", 25),
        ("D1", "Trigger 1: Systematic Profiling", 18),
        ("E1", "Trigger 2: Large-Scale Special Categories", 18),
        ("F1", "Trigger 3: Systematic Monitoring", 18),
        ("G1", "Trigger 4: Innovative Technology", 18),
        ("H1", "Trigger 5: Denial of Service", 18),
        ("I1", "Trigger 6: Large Scale", 18),
        ("J1", "Trigger 7: Matching Datasets", 18),
        ("K1", "Trigger 8: Vulnerable Subjects", 18),
        ("L1", "Trigger 9: Cross-Border Transfer", 18),
        ("M1", "Total Triggers", 12),
        ("N1", "DPIA Required?", 15),
        ("O1", "Notes", 40),
    ]
    
    for cell_ref, header_text, width in headers:
        cell = ws[cell_ref]
        cell.value = header_text
        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = width
    
    style_header_row(ws, 1, COLORS['header_orange'], 15)
    
    # Data validation for trigger columns (D-L: Yes/No)
    for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        add_dropdown(ws, f'{col}2:{col}1000', 'Yes,No', 'Select Yes or No')
    
    # Text validation
    add_dropdown(ws, 'B2:B1000', '', 'Processing name required', allow_blank=True)  # Will be custom validation
    add_dropdown(ws, 'C2:C1000', '', 'System name required', allow_blank=True)
    
    # Formulas
    for row in range(2, 1001):
        # Column M: Total Triggers
        ws[f'M{row}'] = f'=COUNTIF(D{row}:L{row},"Yes")'
        ws[f'M{row}'].number_format = '0'
        ws[f'M{row}'].alignment = Alignment(horizontal='center')
        ws[f'M{row}'].protection = Protection(locked=True)
        
        # Column N: DPIA Required?
        ws[f'N{row}'] = f'=IF(M{row}>=2,"Yes",IF(M{row}=1,"Uncertain","No"))'
        ws[f'N{row}'].alignment = Alignment(horizontal='center')
        ws[f'N{row}'].font = Font(bold=True)
        ws[f'N{row}'].protection = Protection(locked=True)
    
    # Conditional formatting for Total Triggers (Column M)
    ws.conditional_formatting.add('M2:M1000',
        CellIsRule(operator='equal', formula=['0'], 
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid')))
    ws.conditional_formatting.add('M2:M1000',
        CellIsRule(operator='equal', formula=['1'], 
                   fill=PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid')))
    ws.conditional_formatting.add('M2:M1000',
        CellIsRule(operator='greaterThanOrEqual', formula=['2'], 
                   fill=PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid')))
    
    # Conditional formatting for DPIA Required (Column N)
    ws.conditional_formatting.add('N2:N1000',
        CellIsRule(operator='equal', formula=['"No"'], 
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid'),
                   font=Font(color=COLORS['dark_green'])))
    ws.conditional_formatting.add('N2:N1000',
        CellIsRule(operator='equal', formula=['"Uncertain"'], 
                   fill=PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid'),
                   font=Font(color=COLORS['dark_orange'])))
    ws.conditional_formatting.add('N2:N1000',
        CellIsRule(operator='equal', formula=['"Yes"'], 
                   fill=PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid'),
                   font=Font(color=COLORS['dark_red'])))
    
    # Row-level highlighting for DPIA Required = Yes
    ws.conditional_formatting.add('A2:O1000',
        FormulaRule(formula=['$N2="Yes"'], 
                    fill=PatternFill(start_color=COLORS['very_light_red'], end_color=COLORS['very_light_red'], fill_type='solid')))
    
    # Row-level highlighting for DPIA Required = Uncertain
    ws.conditional_formatting.add('A2:O1000',
        FormulaRule(formula=['$N2="Uncertain"'], 
                    fill=PatternFill(start_color=COLORS['very_light_yellow'], end_color=COLORS['very_light_yellow'], fill_type='solid')))
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Protection: Unlock user input cells
    for row in range(2, 1001):
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            if col in [1, 2, 3] or col in range(4, 13) or col == 15:  # A-C, D-L, O
                cell.protection = Protection(locked=False)
            else:  # M, N
                cell.protection = Protection(locked=True)
    
    protect_sheet(ws)


def create_dpia_register_sheet(ws):
    """Create Sheet 3: DPIA Register."""
    
    headers = [
        ("A1", "DPIA ID", 18),
        ("B1", "Processing Activity ID", 20),
        ("C1", "Processing Activity Name", 30),
        ("D1", "System/Application", 25),
        ("E1", "Business Owner", 20),
        ("F1", "DPO Assigned", 20),
        ("G1", "DPIA Start Date", 15),
        ("H1", "Target Completion Date", 15),
        ("I1", "Actual Completion Date", 15),
        ("J1", "DPIA Status", 15),
        ("K1", "Initial Risk Rating", 15),
        ("L1", "Residual Risk Rating", 15),
        ("M1", "Supervisory Authority Consulted?", 18),
        ("N1", "Authority Consultation Date", 15),
        ("O1", "Authority Reference Number", 20),
        ("P1", "Next Review Date", 15),
        ("Q1", "DPIA Document Location", 40),
        ("R1", "Notes", 40),
    ]
    
    for cell_ref, header_text, width in headers:
        cell = ws[cell_ref]
        cell.value = header_text
        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = width
    
    style_header_row(ws, 1, COLORS['header_green'], 18)
    
    # Dropdowns
    add_dropdown(ws, 'E2:E1000', 'HR,IT,Marketing,Sales,Finance,Legal,Operations,Product,Engineering,Other', 
                 'Select department')
    add_dropdown(ws, 'J2:J1000', 'Planned,In Progress,Under Review,Complete,Overdue', 
                 'Select status')
    add_dropdown(ws, 'K2:K1000', 'Low,Medium,High,Critical', 'Select risk level')
    add_dropdown(ws, 'L2:L1000', 'Low,Medium,High,Critical', 'Select risk level')
    add_dropdown(ws, 'M2:M1000', 'Yes,No,N/A', 'Select from list')
    
    # Formulas
    for row in range(2, 1001):
        # Column C: Processing Activity Name (would reference Trigger_Assessment sheet if external link)
        # For standalone workbook, leave empty for manual entry or VLOOKUP to Trigger_Assessment
        ws[f'C{row}'] = f'=IFERROR(VLOOKUP(B{row},Trigger_Assessment!$A$2:$B$1000,2,FALSE),"")'
        ws[f'C{row}'].protection = Protection(locked=True)
        
        # Column D: System/Application
        ws[f'D{row}'] = f'=IFERROR(VLOOKUP(B{row},Trigger_Assessment!$A$2:$C$1000,3,FALSE),"")'
        ws[f'D{row}'].protection = Protection(locked=True)
        
        # Column P: Next Review Date (Completion Date + 1 year)
        ws[f'P{row}'] = f'=IF(ISBLANK(I{row}),"",DATE(YEAR(I{row})+1,MONTH(I{row}),DAY(I{row})))'
        ws[f'P{row}'].number_format = 'DD.MM.YYYY'
        ws[f'P{row}'].protection = Protection(locked=True)
    
    # Date formatting
    for col in ['G', 'H', 'I', 'N', 'P']:
        for row in range(2, 1001):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'
    
    # Conditional formatting for DPIA Status (Column J)
    ws.conditional_formatting.add('J2:J1000',
        CellIsRule(operator='equal', formula=['"Complete"'], 
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid'),
                   font=Font(color=COLORS['dark_green'])))
    ws.conditional_formatting.add('J2:J1000',
        CellIsRule(operator='equal', formula=['"In Progress"'], 
                   fill=PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid'),
                   font=Font(color=COLORS['dark_orange'])))
    ws.conditional_formatting.add('J2:J1000',
        CellIsRule(operator='equal', formula=['"Overdue"'], 
                   fill=PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid'),
                   font=Font(color=COLORS['dark_red'])))
    ws.conditional_formatting.add('J2:J1000',
        CellIsRule(operator='equal', formula=['"Under Review"'], 
                   fill=PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type='solid'),
                   font=Font(color=COLORS['dark_blue'])))
    ws.conditional_formatting.add('J2:J1000',
        CellIsRule(operator='equal', formula=['"Planned"'], 
                   fill=PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type='solid')))
    
    # Risk level formatting
    add_risk_level_formatting(ws, 'K', 2, 1000)
    add_risk_level_formatting(ws, 'L', 2, 1000)
    
    # Next Review Date highlighting (due soon/overdue)
    ws.conditional_formatting.add('P2:P1000',
        FormulaRule(formula=['AND(NOT(ISBLANK(P2)),P2<=TODAY()+30)'], 
                    fill=PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid')))
    ws.conditional_formatting.add('P2:P1000',
        FormulaRule(formula=['AND(NOT(ISBLANK(P2)),P2<=TODAY()+90,P2>TODAY()+30)'], 
                    fill=PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid')))
    
    # Row-level highlighting for overdue DPIAs
    ws.conditional_formatting.add('A2:R1000',
        FormulaRule(formula=['AND($J2="Overdue",OR(NOT(ISBLANK($A2)),NOT(ISBLANK($B2))))'], 
                    fill=PatternFill(start_color=COLORS['very_light_red'], end_color=COLORS['very_light_red'], fill_type='solid'),
                    font=Font(bold=True)))
    
    # Row-level highlighting for complete DPIAs
    ws.conditional_formatting.add('A2:R1000',
        FormulaRule(formula=['$J2="Complete"'], 
                    fill=PatternFill(start_color=COLORS['very_light_green'], end_color=COLORS['very_light_green'], fill_type='solid')))
    
    # Row-level highlighting for high/critical residual risk
    ws.conditional_formatting.add('A2:R1000',
        FormulaRule(formula=['OR($L2="High",$L2="Critical")'], 
                    fill=PatternFill(start_color=COLORS['very_light_orange'], end_color=COLORS['very_light_orange'], fill_type='solid'),
                    border=Border(left=Side(style='thick', color=COLORS['light_red']))))
    
    ws.freeze_panes = 'A2'
    protect_sheet(ws)


def create_risk_assessment_sheet(ws):
    """Create Sheet 4: Risk Assessment Matrix."""
    
    headers = [
        ("A1", "DPIA ID", 18),
        ("B1", "Risk ID", 20),
        ("C1", "Risk Category", 20),
        ("D1", "Risk Description", 50),
        ("E1", "Data Subject Impact", 40),
        ("F1", "Likelihood (Before Mitigation)", 15),
        ("G1", "Impact (Before Mitigation)", 15),
        ("H1", "Inherent Risk Score", 12),
        ("I1", "Inherent Risk Level", 15),
        ("J1", "Necessity Justified?", 15),
        ("K1", "Necessity Justification", 40),
        ("L1", "Proportionality Justified?", 15),
        ("M1", "Proportionality Justification", 40),
        ("N1", "Legal Basis", 20),
        ("O1", "Special Category Legal Basis", 20),
        ("P1", "Data Subject Rights Respected?", 15),
        ("Q1", "Rights Restrictions Documented", 40),
        ("R1", "Third-Party Recipients", 30),
        ("S1", "Cross-Border Transfers", 15),
        ("T1", "Transfer Safeguards", 30),
    ]
    
    for cell_ref, header_text, width in headers:
        cell = ws[cell_ref]
        cell.value = header_text
        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = width
    
    style_header_row(ws, 1, COLORS['header_red'], 20)
    
    # Dropdowns
    risk_categories = "Discrimination,Identity Theft/Fraud,Financial Loss,Reputational Damage,Physical Harm,Loss of Confidentiality,Loss of Control,Surveillance,Profiling with Significant Effects,Social Disadvantage,Psychological Harm,Loss of Rights/Freedoms"
    add_dropdown(ws, 'C2:C1000', risk_categories, 'Select risk category')
    
    likelihood_scale = "1 - Rare,2 - Unlikely,3 - Possible,4 - Likely,5 - Almost Certain"
    add_dropdown(ws, 'F2:F1000', likelihood_scale, 'Select likelihood')
    
    impact_scale = "1 - Negligible,2 - Minor,3 - Moderate,4 - Major,5 - Severe"
    add_dropdown(ws, 'G2:G1000', impact_scale, 'Select impact')
    
    add_dropdown(ws, 'J2:J1000', 'Yes,No,Uncertain', 'Select from list')
    add_dropdown(ws, 'L2:L1000', 'Yes,No,Uncertain', 'Select from list')
    
    legal_basis = "Consent (Art. 6(1)(a)),Contract (Art. 6(1)(b)),Legal Obligation (Art. 6(1)(c)),Vital Interests (Art. 6(1)(d)),Public Task (Art. 6(1)(e)),Legitimate Interests (Art. 6(1)(f))"
    add_dropdown(ws, 'N2:N1000', legal_basis, 'Select GDPR Article 6 legal basis')
    
    special_category_basis = "N/A,Explicit Consent (Art. 9(2)(a)),Employment Law (Art. 9(2)(b)),Vital Interests (Art. 9(2)(c)),Medical Purposes (Art. 9(2)(h)),Public Health (Art. 9(2)(i)),Archiving/Research (Art. 9(2)(j))"
    add_dropdown(ws, 'O2:O1000', special_category_basis, 'Select GDPR Article 9 legal basis or N/A')
    
    add_dropdown(ws, 'P2:P1000', 'Yes,No,Partial', 'Select from list')
    add_dropdown(ws, 'S2:S1000', 'Yes,No', 'Select Yes or No')
    
    # Formulas
    for row in range(2, 1001):
        # Column B: Risk ID (auto-generated)
        ws[f'B{row}'] = f'=IF(ISBLANK(A{row}),"",A{row}&"-R"&TEXT(COUNTIF($A$2:A{row},A{row}),"00"))'
        ws[f'B{row}'].protection = Protection(locked=True)
        
        # Column H: Inherent Risk Score = Likelihood × Impact
        ws[f'H{row}'] = f'=IF(OR(ISBLANK(F{row}),ISBLANK(G{row})),"",VALUE(LEFT(F{row},1))*VALUE(LEFT(G{row},1)))'
        ws[f'H{row}'].number_format = '0'
        ws[f'H{row}'].alignment = Alignment(horizontal='center')
        ws[f'H{row}'].font = Font(bold=True)
        ws[f'H{row}'].protection = Protection(locked=True)
        
        # Column I: Inherent Risk Level
        ws[f'I{row}'] = f'=IF(ISBLANK(H{row}),"",IF(H{row}>=20,"Critical",IF(H{row}>=15,"High",IF(H{row}>=8,"Medium","Low"))))'
        ws[f'I{row}'].alignment = Alignment(horizontal='center')
        ws[f'I{row}'].font = Font(bold=True)
        ws[f'I{row}'].protection = Protection(locked=True)
    
    # Risk score conditional formatting
    for col in ['H', 'I']:
        add_risk_level_formatting(ws, col, 2, 1000)
    
    # Numeric risk score coloring (Column H specific)
    ws.conditional_formatting.add('H2:H1000',
        CellIsRule(operator='between', formula=['1', '6'], 
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid'),
                   font=Font(color=COLORS['dark_green'])))
    ws.conditional_formatting.add('H2:H1000',
        CellIsRule(operator='between', formula=['8', '12'], 
                   fill=PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid'),
                   font=Font(color=COLORS['dark_orange'])))
    ws.conditional_formatting.add('H2:H1000',
        CellIsRule(operator='between', formula=['15', '16'], 
                   fill=PatternFill(start_color=COLORS['light_orange'], end_color=COLORS['light_orange'], fill_type='solid'),
                   font=Font(color=COLORS['dark_red'])))
    ws.conditional_formatting.add('H2:H1000',
        CellIsRule(operator='greaterThanOrEqual', formula=['20'], 
                   fill=PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid'),
                   font=Font(color=COLORS['dark_red'], bold=True)))
    
    # Row-level highlighting for Critical/High risks
    ws.conditional_formatting.add('A2:T1000',
        FormulaRule(formula=['OR($I2="Critical",$I2="High")'], 
                    fill=PatternFill(start_color=COLORS['very_light_red'], end_color=COLORS['very_light_red'], fill_type='solid'),
                    border=Border(left=Side(style='thick', color=COLORS['header_red']))))
    
    # Row-level highlighting for necessity/proportionality failures
    ws.conditional_formatting.add('A2:T1000',
        FormulaRule(formula=['OR($J2="No",$L2="No")'], 
                    fill=PatternFill(start_color=COLORS['very_light_orange'], end_color=COLORS['very_light_orange'], fill_type='solid'),
                    border=Border(right=Side(style='thick', color=COLORS['light_red']))))
    
    ws.freeze_panes = 'A2'
    protect_sheet(ws)


def create_mitigation_measures_sheet(ws):
    """Create Sheet 5: Mitigation Measures."""
    
    headers = [
        ("A1", "Risk ID", 20),
        ("B1", "Risk Description (Reference)", 50),
        ("C1", "Mitigation Control ID", 20),
        ("D1", "Control Type", 20),
        ("E1", "Mitigation Description", 50),
        ("F1", "Implementation Status", 15),
        ("G1", "Owner", 20),
        ("H1", "Target Date", 15),
        ("I1", "Actual Date", 15),
        ("J1", "Effectiveness Rating", 15),
        ("K1", "Evidence Location", 40),
    ]
    
    for cell_ref, header_text, width in headers:
        cell = ws[cell_ref]
        cell.value = header_text
        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = width
    
    style_header_row(ws, 1, COLORS['header_blue'], 11)
    
    # Dropdowns
    add_dropdown(ws, 'D2:D1000', 'Technical,Organizational,Legal,Physical,Administrative', 
                 'Select control type')
    add_dropdown(ws, 'F2:F1000', 'Planned,In Progress,Implemented,Validated,Rejected', 
                 'Select implementation status')
    add_dropdown(ws, 'J2:J1000', 'Not Assessed,Ineffective,Partially Effective,Effective,Highly Effective', 
                 'Select effectiveness rating')
    
    # Formulas
    for row in range(2, 1001):
        # Column B: Risk Description (VLOOKUP)
        ws[f'B{row}'] = f'=IFERROR(VLOOKUP(A{row},Risk_Assessment!$B$2:$D$1000,2,FALSE),"")'
        ws[f'B{row}'].protection = Protection(locked=True)
        
        # Column C: Mitigation Control ID
        ws[f'C{row}'] = f'=IF(ISBLANK(A{row}),"",A{row}&"-M"&TEXT(COUNTIF($A$2:A{row},A{row}),"00"))'
        ws[f'C{row}'].protection = Protection(locked=True)
    
    # Date formatting
    for col in ['H', 'I']:
        for row in range(2, 1001):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'
    
    # Status conditional formatting
    ws.conditional_formatting.add('F2:F1000',
        CellIsRule(operator='equal', formula=['"Validated"'], 
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid'),
                   font=Font(color=COLORS['dark_green'], bold=True)))
    ws.conditional_formatting.add('F2:F1000',
        CellIsRule(operator='equal', formula=['"Implemented"'], 
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid'),
                   font=Font(color=COLORS['dark_green'])))
    ws.conditional_formatting.add('F2:F1000',
        CellIsRule(operator='equal', formula=['"In Progress"'], 
                   fill=PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid'),
                   font=Font(color=COLORS['dark_orange'])))
    ws.conditional_formatting.add('F2:F1000',
        CellIsRule(operator='equal', formula=['"Planned"'], 
                   fill=PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type='solid')))
    ws.conditional_formatting.add('F2:F1000',
        CellIsRule(operator='equal', formula=['"Rejected"'], 
                   fill=PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid'),
                   font=Font(color=COLORS['dark_red'])))
    
    # Effectiveness rating conditional formatting
    ws.conditional_formatting.add('J2:J1000',
        CellIsRule(operator='equal', formula=['"Highly Effective"'], 
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid'),
                   font=Font(color=COLORS['dark_green'])))
    ws.conditional_formatting.add('J2:J1000',
        CellIsRule(operator='equal', formula=['"Effective"'], 
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid'),
                   font=Font(color=COLORS['dark_green'])))
    ws.conditional_formatting.add('J2:J1000',
        CellIsRule(operator='equal', formula=['"Partially Effective"'], 
                   fill=PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid'),
                   font=Font(color=COLORS['dark_orange'])))
    ws.conditional_formatting.add('J2:J1000',
        CellIsRule(operator='equal', formula=['"Ineffective"'], 
                   fill=PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid'),
                   font=Font(color=COLORS['dark_red'])))
    
    # Overdue mitigation highlighting
    ws.conditional_formatting.add('A2:K1000',
        FormulaRule(formula=['AND(NOT(ISBLANK($H2)),$H2<TODAY(),OR($F2="Planned",$F2="In Progress"))'], 
                    fill=PatternFill(start_color=COLORS['very_light_red'], end_color=COLORS['very_light_red'], fill_type='solid'),
                    border=Border(left=Side(style='thick', color=COLORS['header_red']))))
    
    ws.freeze_panes = 'A2'
    protect_sheet(ws)


def create_stakeholder_consultation_sheet(ws):
    """Create Sheet 6: Stakeholder Consultation."""
    
    headers = [
        ("A1", "DPIA ID", 18),
        ("B1", "Stakeholder Type", 20),
        ("C1", "Stakeholder Name/Title", 25),
        ("D1", "Consultation Date", 15),
        ("E1", "Consultation Method", 20),
        ("F1", "Key Concerns Raised", 50),
        ("G1", "Recommendations", 50),
        ("H1", "Action Taken", 50),
        ("I1", "Evidence Location", 40),
    ]
    
    for cell_ref, header_text, width in headers:
        cell = ws[cell_ref]
        cell.value = header_text
        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = width
    
    style_header_row(ws, 1, COLORS['header_green'], 9)
    
    # Dropdowns
    stakeholder_types = "DPO,Data Subjects,Supervisory Authority,Legal Counsel,IT/Security Team,Business Unit,External Consultant,Other"
    add_dropdown(ws, 'B2:B1000', stakeholder_types, 'Select stakeholder type')
    
    consultation_methods = "Meeting,Email,Survey,Workshop,Interview,Public Consultation,Other"
    add_dropdown(ws, 'E2:E1000', consultation_methods, 'Select consultation method')
    
    # Date formatting
    for row in range(2, 1001):
        ws[f'D{row}'].number_format = 'DD.MM.YYYY'
    
    ws.freeze_panes = 'A2'
    protect_sheet(ws)


def create_gap_analysis_sheet(ws):
    """Create Sheet 7: Gap Analysis."""
    
    headers = [
        ("A1", "Risk ID", 20),
        ("B1", "Inherent Risk Score (Reference)", 15),
        ("C1", "Mitigation Implemented?", 15),
        ("D1", "Mitigation Effectiveness", 15),
        ("E1", "Risk Reduction Factor", 12),
        ("F1", "Residual Likelihood", 12),
        ("G1", "Residual Risk Score", 12),
        ("H1", "Residual Risk Level", 15),
        ("I1", "Gap Identified?", 12),
        ("J1", "Gap Description", 50),
        ("K1", "Remediation Plan", 50),
    ]
    
    for cell_ref, header_text, width in headers:
        cell = ws[cell_ref]
        cell.value = header_text
        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = width
    
    style_header_row(ws, 1, COLORS['header_gold'], 11)
    
    # Dropdowns
    add_dropdown(ws, 'C2:C1000', 'Yes,No,Partial', 'Select from list')
    
    effectiveness_options = "N/A,Low (10% reduction),Medium (30% reduction),High (50% reduction),Very High (70% reduction)"
    add_dropdown(ws, 'D2:D1000', effectiveness_options, 'Select effectiveness')
    
    add_dropdown(ws, 'I2:I1000', 'Yes,No', 'Select Yes or No')
    
    # Formulas
    for row in range(2, 1001):
        # Column B: Inherent Risk Score reference
        ws[f'B{row}'] = f'=IFERROR(VLOOKUP(A{row},Risk_Assessment!$B$2:$H$1000,7,FALSE),"")'
        ws[f'B{row}'].number_format = '0'
        ws[f'B{row}'].protection = Protection(locked=True)
        
        # Column E: Risk Reduction Factor
        formula = (
            f'=IF(C{row}="No",1,'
            f'IF(D{row}="N/A",1,'
            f'IF(D{row}="Low (10% reduction)",0.9,'
            f'IF(D{row}="Medium (30% reduction)",0.7,'
            f'IF(D{row}="High (50% reduction)",0.5,'
            f'IF(D{row}="Very High (70% reduction)",0.3,1))))))'
        )
        ws[f'E{row}'] = formula
        ws[f'E{row}'].number_format = '0.00'
        ws[f'E{row}'].protection = Protection(locked=True)
        
        # Column F: Residual Likelihood (approximation)
        ws[f'F{row}'] = f'=IF(ISBLANK(B{row}),"",ROUNDUP(SQRT(B{row})*E{row},0))'
        ws[f'F{row}'].number_format = '0'
        ws[f'F{row}'].protection = Protection(locked=True)
        
        # Column G: Residual Risk Score
        ws[f'G{row}'] = f'=IF(ISBLANK(B{row}),"",ROUNDUP(B{row}*E{row},0))'
        ws[f'G{row}'].number_format = '0'
        ws[f'G{row}'].alignment = Alignment(horizontal='center')
        ws[f'G{row}'].font = Font(bold=True)
        ws[f'G{row}'].protection = Protection(locked=True)
        
        # Column H: Residual Risk Level
        ws[f'H{row}'] = f'=IF(ISBLANK(G{row}),"",IF(G{row}>=20,"Critical",IF(G{row}>=15,"High",IF(G{row}>=8,"Medium","Low"))))'
        ws[f'H{row}'].alignment = Alignment(horizontal='center')
        ws[f'H{row}'].font = Font(bold=True)
        ws[f'H{row}'].protection = Protection(locked=True)
    
    # Risk level formatting
    add_risk_level_formatting(ws, 'G', 2, 1000)
    add_risk_level_formatting(ws, 'H', 2, 1000)
    
    # Gap identification highlighting
    ws.conditional_formatting.add('I2:I1000',
        CellIsRule(operator='equal', formula=['"Yes"'], 
                   fill=PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid'),
                   font=Font(color=COLORS['dark_red'])))
    
    # Row-level highlighting for high/critical residual risk
    ws.conditional_formatting.add('A2:K1000',
        FormulaRule(formula=['OR($H2="High",$H2="Critical")'], 
                    fill=PatternFill(start_color=COLORS['very_light_red'], end_color=COLORS['very_light_red'], fill_type='solid'),
                    border=Border(left=Side(style='thick', color=COLORS['header_red']))))
    
    ws.freeze_panes = 'A2'
    protect_sheet(ws)


def create_dashboard_sheet(ws):
    """Create Sheet 8: Compliance Dashboard (all formulas, locked)."""
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 15
    
    # Section 1: DPIA Summary Metrics
    current_row = 1
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "DPIA ASSESSMENT SUMMARY"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=14, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 2
    
    summary_metrics = [
        ("Total DPIAs Registered", "=COUNTA(DPIA_Register!A2:A1000)", "0"),
        ("DPIAs Complete", '=COUNTIF(DPIA_Register!J2:J1000,"Complete")', "0"),
        ("DPIAs In Progress", '=COUNTIF(DPIA_Register!J2:J1000,"In Progress")', "0"),
        ("DPIAs Overdue", '=COUNTIF(DPIA_Register!J2:J1000,"Overdue")', "0"),
        ("Completion Rate", "=IF(C3=0,0,C4/C3)", "0%"),
        ("", "", ""),
        ("Supervisory Authority Consultations", '=COUNTIF(DPIA_Register!M2:M1000,"Yes")', "0"),
    ]
    
    for label, formula, num_format in summary_metrics:
        ws[f'A{current_row}'].value = label
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'C{current_row}'] = formula
        ws[f'C{current_row}'].number_format = num_format
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
    
    # Completion Rate conditional formatting
    completion_rate_cell = 'C7'
    ws.conditional_formatting.add(completion_rate_cell,
        CellIsRule(operator='greaterThanOrEqual', formula=['0.8'], 
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid'),
                   font=Font(color=COLORS['dark_green'], bold=True)))
    ws.conditional_formatting.add(completion_rate_cell,
        CellIsRule(operator='between', formula=['0.5', '0.8'], 
                   fill=PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid'),
                   font=Font(color=COLORS['dark_orange'], bold=True)))
    ws.conditional_formatting.add(completion_rate_cell,
        CellIsRule(operator='lessThan', formula=['0.5'], 
                   fill=PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid'),
                   font=Font(color=COLORS['dark_red'], bold=True)))
    
    # Section 2: Risk Distribution
    current_row += 2
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "RISK DISTRIBUTION"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=12, bold=True)
    current_row += 1
    
    risk_metrics = [
        ("Total Risks Identified", "=COUNTA(Risk_Assessment!B2:B1000)", "0"),
        ("Critical Risks", '=COUNTIF(Risk_Assessment!I2:I1000,"Critical")', "0"),
        ("High Risks", '=COUNTIF(Risk_Assessment!I2:I1000,"High")', "0"),
        ("Medium Risks", '=COUNTIF(Risk_Assessment!I2:I1000,"Medium")', "0"),
        ("Low Risks", '=COUNTIF(Risk_Assessment!I2:I1000,"Low")', "0"),
    ]
    
    for label, formula, num_format in risk_metrics:
        ws[f'A{current_row}'].value = label
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'C{current_row}'] = formula
        ws[f'C{current_row}'].number_format = num_format
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
    
    # Section 3: Mitigation Implementation
    current_row += 2
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "MITIGATION IMPLEMENTATION"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=12, bold=True)
    current_row += 1
    
    mitigation_metrics = [
        ("Total Mitigation Controls", "=COUNTA(Mitigation_Measures!C2:C1000)", "0"),
        ("Controls Validated", '=COUNTIF(Mitigation_Measures!F2:F1000,"Validated")', "0"),
        ("Controls Implemented", '=COUNTIF(Mitigation_Measures!F2:F1000,"Implemented")', "0"),
        ("Controls In Progress", '=COUNTIF(Mitigation_Measures!F2:F1000,"In Progress")', "0"),
        ("Mitigation Completion Rate", "=IF(C24=0,0,(C25+C26)/C24)", "0%"),
    ]
    
    for label, formula, num_format in mitigation_metrics:
        ws[f'A{current_row}'].value = label
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'C{current_row}'] = formula
        ws[f'C{current_row}'].number_format = num_format
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
    
    # Mitigation completion rate conditional formatting
    mitigation_rate_cell = 'C28'
    ws.conditional_formatting.add(mitigation_rate_cell,
        CellIsRule(operator='greaterThanOrEqual', formula=['0.8'], 
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid'),
                   font=Font(color=COLORS['dark_green'], bold=True)))
    ws.conditional_formatting.add(mitigation_rate_cell,
        CellIsRule(operator='between', formula=['0.5', '0.8'], 
                   fill=PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid'),
                   font=Font(color=COLORS['dark_orange'], bold=True)))
    ws.conditional_formatting.add(mitigation_rate_cell,
        CellIsRule(operator='lessThan', formula=['0.5'], 
                   fill=PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid'),
                   font=Font(color=COLORS['dark_red'], bold=True)))
    
    # Section 4: Residual Risk Summary
    current_row += 2
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "RESIDUAL RISK AFTER MITIGATION"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=12, bold=True)
    current_row += 1
    
    residual_metrics = [
        ("Total Assessed Risks", "=COUNTA(Gap_Analysis!A2:A1000)", "0"),
        ("Residual Risk - Critical", '=COUNTIF(Gap_Analysis!H2:H1000,"Critical")', "0"),
        ("Residual Risk - High", '=COUNTIF(Gap_Analysis!H2:H1000,"High")', "0"),
        ("Residual Risk - Medium", '=COUNTIF(Gap_Analysis!H2:H1000,"Medium")', "0"),
        ("Residual Risk - Low", '=COUNTIF(Gap_Analysis!H2:H1000,"Low")', "0"),
    ]
    
    for label, formula, num_format in residual_metrics:
        ws[f'A{current_row}'].value = label
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'C{current_row}'] = formula
        ws[f'C{current_row}'].number_format = num_format
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
    
    # Section 5: GDPR Compliance Indicators
    current_row += 2
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "GDPR ARTICLE 35 COMPLIANCE"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=12, bold=True)
    current_row += 1
    
    compliance_metrics = [
        ("Necessity Justified (Yes)", '=COUNTIF(Risk_Assessment!J2:J1000,"Yes")', "0"),
        ("Proportionality Justified (Yes)", '=COUNTIF(Risk_Assessment!L2:L1000,"Yes")', "0"),
        ("Data Subject Rights Fully Respected", '=COUNTIF(Risk_Assessment!P2:P1000,"Yes")', "0"),
        ("Stakeholder Consultations Conducted", "=COUNTA(Stakeholder_Consultation!A2:A1000)", "0"),
        ("", "", ""),
        ("Overall DPIA Compliance Score", "=IF(C3=0,0,C7*0.4+C28*0.3+IF(C34=0,0,(1-C36/C34)*0.3))", "0%"),
    ]
    
    for label, formula, num_format in compliance_metrics:
        ws[f'A{current_row}'].value = label
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'C{current_row}'] = formula
        ws[f'C{current_row}'].number_format = num_format
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
    
    # Overall compliance score conditional formatting
    compliance_score_cell = 'C49'
    ws.conditional_formatting.add(compliance_score_cell,
        CellIsRule(operator='greaterThanOrEqual', formula=['0.8'], 
                   fill=PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid'),
                   font=Font(color=COLORS['dark_green'], bold=True, size=12)))
    ws.conditional_formatting.add(compliance_score_cell,
        CellIsRule(operator='between', formula=['0.6', '0.8'], 
                   fill=PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid'),
                   font=Font(color=COLORS['dark_orange'], bold=True, size=12)))
    ws.conditional_formatting.add(compliance_score_cell,
        CellIsRule(operator='lessThan', formula=['0.6'], 
                   fill=PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid'),
                   font=Font(color=COLORS['dark_red'], bold=True, size=12)))
    
    # Add charts (optional, simplified for script length)
    # Chart 1: DPIA Status Pie Chart would go here
    # Chart 2: Risk Distribution Bar Chart would go here
    # Chart 3: Residual vs Inherent Risk comparison would go here
    
    # Lock entire dashboard sheet
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)
    
    protect_sheet(ws)


def main():
    """Main function to generate DPIA assessment workbook."""
    
    logger.info("=" * 80)
    logger.info("ISMS A.5.34.5 - DPIA Assessment Workbook Generator")
    logger.info("=" * 80)
    logger.info("")
    
    try:
        # Create workbook
        logger.info("Creating workbook...")
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create 8 sheets
        logger.info("Creating Sheet 1: Instructions & Legend...")
        ws1 = wb.create_sheet("Instructions", 0)
        ws1.sheet_properties.tabColor = COLORS['header_blue'].replace('FF', '')
        create_instructions_sheet(ws1)
        
        logger.info("Creating Sheet 2: Trigger Assessment...")
        ws2 = wb.create_sheet("Trigger_Assessment", 1)
        ws2.sheet_properties.tabColor = COLORS['header_orange'].replace('FF', '')
        create_trigger_assessment_sheet(ws2)
        
        logger.info("Creating Sheet 3: DPIA Register...")
        ws3 = wb.create_sheet("DPIA_Register", 2)
        ws3.sheet_properties.tabColor = COLORS['header_green'].replace('FF', '')
        create_dpia_register_sheet(ws3)
        
        logger.info("Creating Sheet 4: Risk Assessment...")
        ws4 = wb.create_sheet("Risk_Assessment", 3)
        ws4.sheet_properties.tabColor = COLORS['header_red'].replace('FF', '')
        create_risk_assessment_sheet(ws4)
        
        logger.info("Creating Sheet 5: Mitigation Measures...")
        ws5 = wb.create_sheet("Mitigation_Measures", 4)
        ws5.sheet_properties.tabColor = COLORS['header_blue'].replace('FF', '')
        create_mitigation_measures_sheet(ws5)
        
        logger.info("Creating Sheet 6: Stakeholder Consultation...")
        ws6 = wb.create_sheet("Stakeholder_Consultation", 5)
        ws6.sheet_properties.tabColor = COLORS['header_green'].replace('FF', '')
        create_stakeholder_consultation_sheet(ws6)
        
        logger.info("Creating Sheet 7: Gap Analysis...")
        ws7 = wb.create_sheet("Gap_Analysis", 6)
        ws7.sheet_properties.tabColor = COLORS['header_gold'].replace('FF', '')
        create_gap_analysis_sheet(ws7)
        
        logger.info("Creating Sheet 8: Dashboard...")
        ws8 = wb.create_sheet("Dashboard", 7)
        ws8.sheet_properties.tabColor = COLORS['header_cyan'].replace('FF', '')
        create_dashboard_sheet(ws8)
        
        # Set workbook properties
        wb.properties.title = "A.5.34.5 DPIA Assessment"
        wb.properties.subject = "Data Protection Impact Assessment"
        wb.properties.creator = "ISMS Automation Script"
        wb.properties.created = datetime.now()
        wb.properties.category = "Privacy Compliance"
        wb.properties.keywords = "GDPR, DPIA, Privacy, ISO 27001, A.5.34"
        wb.properties.comments = "Generated workbook for conducting DPIAs per GDPR Article 35"
        
        # Save workbook
        filename = f"ISMS-IMP-A.5.34.5_DPIA_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
        logger.info(f"\nSaving workbook: {filename}")
        wb.save(filename)
        
        logger.info("")
        logger.info("=" * 80)
        logger.info("✅ WORKBOOK CREATED SUCCESSFULLY")
        logger.info("=" * 80)
        logger.info(f"Filename: {filename}")
        logger.info(f"Sheets: 8 (Instructions, Trigger Assessment, DPIA Register, Risk Assessment,")
        logger.info(f"           Mitigation Measures, Stakeholder Consultation, Gap Analysis, Dashboard)")
        logger.info("")
        logger.info("Next Steps:")
        logger.info("1. Open workbook in Excel/LibreOffice")
        logger.info("2. Read Instructions sheet (Sheet 1)")
        logger.info("3. Complete Trigger Assessment (Sheet 2) to identify DPIAs required")
        logger.info("4. For each DPIA required, complete Sheets 3-7 sequentially")
        logger.info("5. Review Dashboard (Sheet 8) for compliance metrics")
        logger.info("")
        logger.info("Compliance Framework: GDPR Article 35, ISO/IEC 27001:2022 A.5.34")
        logger.info("=" * 80)
        
    except Exception as e:
        logger.error(f"\n❌ ERROR: Failed to create workbook")
        logger.error(f"Error details: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
