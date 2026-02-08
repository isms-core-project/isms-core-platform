#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.34.1 - PII Identification and Classification Assessment Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.34: Privacy and Protection of PII
Assessment Domain 1 of 7: PII Identification, Classification, and ROPA

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific PII processing activities, data systems, and
privacy compliance requirements.

Key customization areas:
1. System types and PII categories (match your actual infrastructure)
2. Legal basis options (adapt to applicable regulations - GDPR, FADP, others)
3. Cross-border transfer mechanisms (specific to your data flows)
4. Dropdown lists and validation rules (align with your governance)
5. Compliance thresholds (based on your regulatory requirements)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for privacy)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for identifying
and classifying all personally identifiable information (PII) processed by the
organization, maintaining GDPR Article 30 compliant Record of Processing Activities.

**Purpose:**
Enables systematic PII discovery, classification, and data flow mapping to
demonstrate ISO 27001:2022 Control A.5.34 compliance and meet GDPR/FADP
regulatory requirements.

**Assessment Scope:**
- PII-containing systems inventory (CRM, HR, email, databases, SaaS, etc.)
- PII classification (Basic PII, Sensitive PII, Criminal Offense Data)
- Data flow mapping (collection, processing, storage, sharing, deletion)
- Cross-border transfer identification and safeguard verification
- Record of Processing Activities (ROPA) per GDPR Article 30
- Legal basis documentation for all processing activities
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and GDPR/FADP requirements
2. PII System Inventory - All systems processing PII with classifications
3. PII Data Flow Mapping - Data flows across organizational boundaries
4. Record of Processing Activities (ROPA) - GDPR Art. 30 compliant entries
5. PII Discovery Gaps - Non-compliant processing and remediation plans
6. Evidence Register - Audit evidence tracking and documentation
7. Dashboard - Consolidated compliance metrics and status
8. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with GDPR/FADP legal basis dropdown lists
- Conditional formatting for PII classification and risk levels
- Automated gap identification for missing legal bases
- ROPA structure fully compliant with GDPR Article 30 requirements
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Dashboard with compliance metrics and trend analysis

**Integration:**
This assessment is Domain 1 of the A.5.34 Privacy assessment suite (7 domains).
Results feed into ISMS-IMP-A.5.34.7 Privacy Compliance Dashboard for
consolidated privacy program oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a5341_pii_identification_assessment.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a5341_pii_identification_assessment.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a5341_pii_identification_assessment.py --date 20250128

Output:
    File: ISMS_A_5_34_1_PII_Identification_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize PII categories to match your processing
    2. Inventory all systems processing PII (automated + manual discovery)
    3. Complete PII classification for each system
    4. Map data flows including cross-border transfers
    5. Create ROPA entries for all processing activities
    6. Document legal basis for each processing activity
    7. Conduct gap analysis for unlawful or undocumented processing
    8. Collect and link audit evidence
    9. Review dashboard metrics
    10. Obtain stakeholder approvals
    11. Feed results into A.5.34.7 Privacy Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.34
Assessment Domain:    1 of 7 (PII Identification and Classification)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.34: Privacy and Protection of PII Policy (Governance)
    - ISMS-CTX-A.5.34: Privacy Regulatory Landscape Reference (Context)
    - ISMS-IMP-A.5.34.1: PII Identification Implementation Guide (Part 1 & 2)
    - ISMS-IMP-A.5.34.2: Legal Basis Assessment (Domain 2)
    - ISMS-IMP-A.5.34.3: Data Subject Rights Management (Domain 3)
    - ISMS-IMP-A.5.34.4: Technical and Organizational Measures (Domain 4)
    - ISMS-IMP-A.5.34.5: Data Protection Impact Assessment (Domain 5)
    - ISMS-IMP-A.5.34.6: Cross-Border Transfer Assessment (Domain 6)
    - ISMS-IMP-A.5.34.7: Privacy Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 2025-01-28
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.34.1 specification
    - Supports comprehensive PII identification and ROPA maintenance
    - Integrated dashboard for compliance monitoring
    - Prepares for consolidation into A.5.34.7 Privacy Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Privacy Regulations:**
Privacy regulations evolve rapidly. Review GDPR/FADP guidance from EDPB and
FDPIC quarterly. Update assessment criteria for new regulatory requirements
(DPIAs, consent management, cross-border transfers).

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 and GDPR requirements.
Auditors will expect complete ROPA, documented legal bases, and evidence of
cross-border transfer safeguards (SCCs, BCRs, adequacy decisions).

**Data Protection:**
Assessment workbooks contain sensitive information including:
- Complete inventory of PII-processing systems
- Data flow maps revealing organizational structure
- Gap analysis highlighting compliance deficiencies
- ROPA with detailed processing activity descriptions

Handle in accordance with your organization's data classification policies.
Restrict access to DPO, Legal, CISO, and authorized privacy team members.

**Maintenance:**
Review and update assessment:
- Quarterly: New systems, changed processing, updated ROPA
- Annually: Complete reassessment of all systems and data flows
- Triggered: M&A activity, new products/services, regulatory changes
- Continuous: New processor agreements, cross-border transfer changes

**Quality Assurance:**
Have DPO/Privacy Officer and Legal counsel validate assessments before using
results for compliance reporting, regulatory filings, or audit evidence.

**Regulatory Alignment:**
This assessment supports compliance with:
- GDPR (EU General Data Protection Regulation) - Art. 5, 6, 9, 30
- FADP (Swiss Federal Act on Data Protection) - Art. 5, 6, 8
- ISO/IEC 27001:2022 - Control A.5.34
- ISO/IEC 27701:2019 - Privacy Information Management (if applicable)

================================================================================
"""

import argparse
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.34.1"
WORKBOOK_NAME = "PII Identification and Classification"
CONTROL_ID = "A.5.34"
CONTROL_NAME = "Privacy and Protection of PII"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# CONFIGURATION CONSTANTS
# ============================================================================

# CUSTOMIZE: Modify these lists based on your organization's specifics

SYSTEM_TYPES = [
    "Customer Relationship Management (CRM)",
    "Human Resources Information System (HRIS)",
    "Payroll System",
    "Email System",
    "Collaboration Tool (Slack, Teams, etc.)",
    "File Storage (Google Drive, SharePoint, etc.)",
    "Marketing Automation",
    "Analytics Platform",
    "E-Commerce / Shopping Cart",
    "Payment Processing",
    "Customer Support / Help Desk",
    "Backup System",
    "Database Server",
    "Website / Web Application",
    "Mobile Application",
    "CCTV / Physical Security",
    "Access Control System",
    "Other"
]

PII_DATA_SUBJECTS = [
    "Customers",
    "Employees",
    "Contractors",
    "Vendors/Suppliers",
    "Job Applicants",
    "Website Visitors",
    "Other"
]

PII_CLASSIFICATIONS = [
    "Basic PII",
    "Sensitive PII (GDPR Art. 9 / FADP Art. 5(c))",
    "Criminal Offense Data (GDPR Art. 10 / FADP Art. 5(c)5)"
]

SENSITIVE_PII_TYPES = [
    "Health Data",
    "Genetic Data",
    "Biometric Data (for unique identification)",
    "Racial or Ethnic Origin",
    "Political Opinions",
    "Religious or Philosophical Beliefs",
    "Trade Union Membership",
    "Sex Life or Sexual Orientation",
    "Private Sphere (FADP-specific)"
]

DISCOVERY_METHODS = [
    "Automated Scan (DLP)",
    "Manual Survey",
    "Interview (System Owner)",
    "Documentation Review",
    "Database Schema Analysis",
    "Other"
]

LEGAL_BASIS_ART6 = [
    "Consent (Art. 6(1)(a))",
    "Contract (Art. 6(1)(b))",
    "Legal Obligation (Art. 6(1)(c))",
    "Vital Interests (Art. 6(1)(d))",
    "Public Task (Art. 6(1)(e))",
    "Legitimate Interests (Art. 6(1)(f))"
]

LEGAL_BASIS_ART9 = [
    "N/A (No special category data)",
    "Explicit Consent (Art. 9(2)(a))",
    "Employment / Social Security (Art. 9(2)(b))",
    "Vital Interests (Art. 9(2)(c))",
    "Legitimate Activities of Foundation/Association (Art. 9(2)(d))",
    "Made Public by Data Subject (Art. 9(2)(e))",
    "Legal Claims (Art. 9(2)(f))",
    "Substantial Public Interest (Art. 9(2)(g))",
    "Health / Social Care (Art. 9(2)(h))",
    "Public Health (Art. 9(2)(i))",
    "Archiving / Research / Statistics (Art. 9(2)(j))"
]

TRANSFER_MECHANISMS = [
    "No cross-border transfer",
    "Adequacy Decision (EU/CH recognizes country)",
    "Standard Contractual Clauses (SCCs)",
    "Binding Corporate Rules (BCRs)",
    "EU-US Data Privacy Framework (DPF) Certification",
    "Derogation - Explicit Consent",
    "Derogation - Contract Necessity",
    "Derogation - Legal Claims",
    "Derogation - Vital Interests",
    "Derogation - Legitimate Interests (Occasional)"
]

RECIPIENT_TYPES = [
    "Internal (within organization)",
    "Processor (processes on our behalf per Art. 28)",
    "Third Party (separate controller)",
    "Public Authority (government, regulator)"
]

GAP_TYPES = [
    "Unknown System (Not in inventory)",
    "Undocumented Data Flow",
    "Missing Legal Basis",
    "Unclear Retention Period",
    "Unverified Cross-Border Transfer",
    "Missing Processor Agreement (GDPR Art. 28)",
    "Incomplete PII Categories",
    "Missing DPIA (High-risk processing)",
    "Missing Consent Records",
    "Shadow IT (Unknown to IT)",
    "Unlawful Processing",
    "Inadequate Security Measures",
    "Other"
]

RISK_LEVELS = [
    "Critical",
    "High",
    "Medium",
    "Low"
]

EVIDENCE_TYPES = [
    "System Documentation",
    "Interview Notes",
    "DLP Scan Results",
    "Contract (Processor Agreement, SCCs)",
    "Consent Form/Mechanism",
    "Data Flow Diagram",
    "Legal Basis Assessment (LIA)",
    "DPIA (Data Protection Impact Assessment)",
    "Retention Policy",
    "Screenshot (System interface)",
    "Email Confirmation",
    "Audit Report",
    "Privacy Policy/Notice",
    "Training Records",
    "Other"
]

STATUS_OPTIONS = [
    "Not Started",
    "In Progress",
    "Complete",
    "Validated"
]

SIGNATORY_ROLES = [
    "Assessment Lead (DPO / Privacy Officer)",
    "Chief Information Security Officer (CISO)",
    "Legal / Compliance Officer",
    "Data Owner - HR",
    "Data Owner - Sales/Marketing",
    "Data Owner - Finance",
    "Data Owner - IT",
    "Executive Sponsor"
]

# Color scheme
COLOR_HEADER = "003366"  # Dark blue
COLOR_NOT_STARTED = "D3D3D3"  # Gray
COLOR_IN_PROGRESS = "FFFFCC"  # Yellow
COLOR_COMPLETE = "C6EFCE"  # Light green
COLOR_VALIDATED = "006400"  # Dark green
COLOR_CRITICAL = "8B0000"  # Dark red
COLOR_HIGH = "FFA500"  # Orange
COLOR_MEDIUM = "FFFFCC"  # Yellow
COLOR_LOW = "87CEEB"  # Light blue
COLOR_SENSITIVE_PII = "FFA500"  # Orange


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def create_header_row(ws, headers, start_row=1):
    """Create formatted header row with styling"""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    return ws


def add_data_validation(ws, cell_range, options, allow_blank=True):
    """Add dropdown data validation to cell range"""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=allow_blank,
        showDropDown=True
    )
    dv.error = 'Invalid value. Please select from dropdown.'
    dv.errorTitle = 'Invalid Entry'
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return ws


def add_status_conditional_formatting(ws, status_column, start_row, end_row):
    """Add conditional formatting for status column"""
    col_letter = get_column_letter(status_column)
    
    # Not Started = Gray
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"Not Started"'],
                   fill=PatternFill(start_color=COLOR_NOT_STARTED, end_color=COLOR_NOT_STARTED, fill_type="solid"))
    )
    
    # In Progress = Yellow
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"In Progress"'],
                   fill=PatternFill(start_color=COLOR_IN_PROGRESS, end_color=COLOR_IN_PROGRESS, fill_type="solid"))
    )
    
    # Complete = Light Green
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"Complete"'],
                   fill=PatternFill(start_color=COLOR_COMPLETE, end_color=COLOR_COMPLETE, fill_type="solid"))
    )
    
    # Validated = Dark Green with white text
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"Validated"'],
                   fill=PatternFill(start_color=COLOR_VALIDATED, end_color=COLOR_VALIDATED, fill_type="solid"),
                   font=Font(color="FFFFFF", bold=True))
    )
    
    return ws


def add_risk_conditional_formatting(ws, risk_column, start_row, end_row):
    """Add conditional formatting for risk level column"""
    col_letter = get_column_letter(risk_column)
    
    # Critical = Dark Red with white text
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"Critical"'],
                   fill=PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type="solid"),
                   font=Font(color="FFFFFF", bold=True))
    )
    
    # High = Orange
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"High"'],
                   fill=PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type="solid"))
    )
    
    # Medium = Yellow
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"Medium"'],
                   fill=PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type="solid"))
    )
    
    # Low = Light Blue
    ws.conditional_formatting.add(
        f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='equal', formula=['"Low"'],
                   fill=PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type="solid"))
    )
    
    return ws


def set_column_widths(ws, column_widths):
    """Set column widths based on dictionary {column_number: width}"""
    for col_num, width in column_widths.items():
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = width
    return ws


def protect_sheet(ws, allow_edit_ranges=None):
    """Protect sheet with password, allowing edits to specific ranges"""
    ws.protection.sheet = True
    ws.protection.password = "privacy2024"  # CUSTOMIZE: Change password
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertColumns = False
    ws.protection.insertRows = False
    ws.protection.insertHyperlinks = True
    ws.protection.deleteColumns = False
    ws.protection.deleteRows = False
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True
    ws.protection.sort = False
    ws.protection.autoFilter = False
    ws.protection.pivotTables = False
    ws.protection.objects = True
    ws.protection.scenarios = True
    
    # Unlock input cells if ranges specified
    if allow_edit_ranges:
        for cell_range in allow_edit_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    cell.protection = Protection(locked=False)
    
    return ws


# ============================================================================
# SHEET CREATION FUNCTIONS
# ============================================================================

def create_instructions_sheet(wb):
    """Create Sheet 1: Instructions & Legend"""
    ws = wb.create_sheet("1. Instructions & Legend", 0)
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "ISMS-IMP-A.5.34.1 - PII Identification and Classification Assessment"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Instructions
    instructions = [
        ("", ""),
        ("Purpose:", "This workbook supports systematic identification and classification of all PII processed by the organization."),
        ("", ""),
        ("Regulatory Compliance:", "- GDPR Article 30 (Record of Processing Activities - ROPA)"),
        ("", "- Swiss FADP Articles 5, 6, 8 (Data protection principles, PII categories)"),
        ("", "- ISO/IEC 27001:2022 Control A.5.34 (Privacy and Protection of PII)"),
        ("", ""),
        ("Assessment Scope:", "All systems, processes, and activities that collect, process, store, or transmit personal data."),
        ("", ""),
        ("Completion Instructions:", ""),
        ("", "1. Complete Sheet 2 (PII System Inventory) - identify all PII-containing systems"),
        ("", "2. Complete Sheet 3 (Data Flow Mapping) - map PII flows including cross-border transfers"),
        ("", "3. Complete Sheet 4 (ROPA) - create GDPR-compliant processing activity records"),
        ("", "4. Complete Sheet 5 (Gaps) - document non-compliant processing and remediation plans"),
        ("", "5. Complete Sheet 6 (Evidence) - link supporting evidence for audit trail"),
        ("", "6. Review Sheet 7 (Dashboard) - validate compliance metrics"),
        ("", "7. Obtain approvals in Sheet 8 (Sign-Off)"),
        ("", ""),
        ("Key Definitions:", ""),
        ("", ""),
        ("Personal Data / PII:", "Any information relating to an identified or identifiable natural person (GDPR Art. 4(1), FADP Art. 5(a))"),
        ("", ""),
        ("PII Classifications:", ""),
        ("Basic PII:", "Standard personal data requiring baseline protection (names, emails, addresses, financial data)"),
        ("Sensitive PII:", "Special category data per GDPR Art. 9 / FADP Art. 5(c) requiring enhanced protection:"),
        ("", "  • Health data, genetic data, biometric data (for unique identification)"),
        ("", "  • Racial/ethnic origin, political opinions, religious/philosophical beliefs"),
        ("", "  • Trade union membership, sex life/sexual orientation, private sphere (FADP)"),
        ("Criminal Offense Data:", "Data relating to criminal convictions, offenses, or related security measures (GDPR Art. 10)"),
        ("", ""),
        ("Legal Bases (GDPR Art. 6):", ""),
        ("", "Every processing activity MUST have a valid legal basis:"),
        ("Consent:", "Data subject gave explicit consent for specific purpose (freely given, specific, informed, unambiguous, withdrawable)"),
        ("Contract:", "Processing necessary to perform contract with data subject"),
        ("Legal Obligation:", "Processing required by EU/Member State law"),
        ("Vital Interests:", "Processing necessary to protect life/physical integrity (emergency only)"),
        ("Public Task:", "Processing for task in public interest (typically public authorities only)"),
        ("Legitimate Interests:", "Processing necessary for legitimate interests of controller/third party (requires LIA)"),
        ("", ""),
        ("Cross-Border Transfers:", "Personal data transferred outside EU/Switzerland requires safeguards:"),
        ("Adequacy Decision:", "EU/CH recognizes destination country as providing adequate protection"),
        ("SCCs:", "Standard Contractual Clauses (EU Commission approved contract terms)"),
        ("BCRs:", "Binding Corporate Rules (intra-group transfer mechanism)"),
        ("Derogations:", "Specific exceptions per GDPR Art. 49 / FADP Art. 17 (consent, contract, legal claims)"),
        ("", ""),
        ("Record of Processing Activities (ROPA):", "GDPR Article 30 mandates controllers maintain ROPA documenting:"),
        ("", "  • Purposes of processing"),
        ("", "  • Categories of data subjects and personal data"),
        ("", "  • Categories of recipients (including international transfers)"),
        ("", "  • Retention periods"),
        ("", "  • Technical and organizational security measures"),
        ("", ""),
        ("Assessment Frequency:", ""),
        ("Initial:", "Comprehensive assessment of all systems and processing activities"),
        ("Quarterly:", "Update for new systems, changed processing, M&A activity"),
        ("Annual:", "Complete re-validation of all entries"),
        ("Triggered:", "Material changes (new products, major system changes, regulatory changes)"),
        ("", ""),
        ("Support & Questions:", "Contact DPO/Privacy Officer: [dpo@organization.com]"),
        ("", ""),
        ("Document Version:", "1.0"),
        ("Generated Date:", datetime.now().strftime("%d.%m.%Y")),
    ]
    
    for row_num, (label, text) in enumerate(instructions, 3):
        ws.cell(row=row_num, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_num, column=2, value=text).alignment = Alignment(wrap_text=True, vertical="top")
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 100
    
    return ws


def create_system_inventory_sheet(wb):
    """Create Sheet 2: PII System Inventory"""
    ws = wb.create_sheet("2. PII System Inventory")
    
    headers = [
        "RowID",
        "Status",
        "System Name",
        "System Owner",
        "System Type",
        "PII Processing Role",
        "PII Data Subjects",
        "PII Categories",
        "PII Classification",
        "Sensitive PII Types (if applicable)",
        "Data Volume (approx. records)",
        "Hosting Location",
        "Access Level (who/how many)",
        "Discovery Method",
        "Discovery Date",
        "Last Updated",
        "Updated By",
        "Notes",
        "Evidence Reference",
        "Review Comments"
    ]
    
    create_header_row(ws, headers)
    
    # Add data validation
    add_data_validation(ws, "B2:B1000", STATUS_OPTIONS)
    add_data_validation(ws, "E2:E1000", SYSTEM_TYPES)
    add_data_validation(ws, "F2:F1000", ["Controller", "Processor", "Joint Controller"])
    add_data_validation(ws, "I2:I1000", PII_CLASSIFICATIONS)
    add_data_validation(ws, "N2:N1000", DISCOVERY_METHODS)
    
    # Add conditional formatting
    add_status_conditional_formatting(ws, 2, 2, 1000)
    
    # Highlight Sensitive/Criminal PII
    ws.conditional_formatting.add(
        'I2:I1000',
        CellIsRule(operator='containsText', formula=['"Sensitive"'],
                   fill=PatternFill(start_color=COLOR_SENSITIVE_PII, end_color=COLOR_SENSITIVE_PII, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'I2:I1000',
        CellIsRule(operator='containsText', formula=['"Criminal"'],
                   fill=PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type="solid"),
                   font=Font(color="FFFFFF", bold=True))
    )
    
    # Set column widths
    column_widths = {
        1: 12,  # RowID
        2: 12,  # Status
        3: 25,  # System Name
        4: 25,  # System Owner
        5: 30,  # System Type
        6: 18,  # PII Processing Role
        7: 20,  # PII Data Subjects
        8: 35,  # PII Categories
        9: 20,  # PII Classification
        10: 30, # Sensitive PII Types
        11: 18, # Data Volume
        12: 30, # Hosting Location
        13: 25, # Access Level
        14: 20, # Discovery Method
        15: 15, # Discovery Date
        16: 15, # Last Updated
        17: 20, # Updated By
        18: 35, # Notes
        19: 20, # Evidence Reference
        20: 25  # Review Comments
    }
    set_column_widths(ws, column_widths)
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Add sample row with formulas
    ws['A2'] = '=TEXT(ROW()-1,"SYS-000")'
    ws['O2'] = '=TODAY()'
    ws['P2'] = '=TODAY()'
    
    # Protect sheet, allow edits to data cells
    protect_sheet(ws, allow_edit_ranges=['B2:T1000'])
    
    return ws


def create_data_flow_mapping_sheet(wb):
    """Create Sheet 3: PII Data Flow Mapping"""
    ws = wb.create_sheet("3. PII Data Flow Mapping")
    
    headers = [
        "RowID",
        "Status",
        "Source System",
        "Destination System",
        "PII Categories Transferred",
        "Transfer Method",
        "Transfer Frequency",
        "Purpose of Transfer",
        "Legal Basis (Art. 6)",
        "Recipient Type",
        "Recipient Name (if external)",
        "Cross-Border Transfer?",
        "Destination Country",
        "Transfer Mechanism",
        "SCC Date (if applicable)",
        "TIA Completed? (if cross-border)",
        "Security Measures",
        "Last Updated",
        "Updated By",
        "Notes",
        "Evidence Reference",
        "Review Comments"
    ]
    
    create_header_row(ws, headers)
    
    # Add data validation
    add_data_validation(ws, "B2:B1000", STATUS_OPTIONS)
    add_data_validation(ws, "F2:F1000", ["API", "File Transfer", "Email", "Manual Entry", "Database Replication", "Backup", "Other"])
    add_data_validation(ws, "G2:G1000", ["Real-time", "Hourly", "Daily", "Weekly", "Monthly", "On-Demand", "One-Time"])
    add_data_validation(ws, "I2:I1000", LEGAL_BASIS_ART6)
    add_data_validation(ws, "J2:J1000", RECIPIENT_TYPES)
    add_data_validation(ws, "L2:L1000", ["Yes", "No"])
    add_data_validation(ws, "N2:N1000", TRANSFER_MECHANISMS)
    add_data_validation(ws, "P2:P1000", ["Yes", "No", "N/A"])
    
    # Add conditional formatting
    add_status_conditional_formatting(ws, 2, 2, 1000)
    
    # Highlight cross-border transfers without mechanism
    ws.conditional_formatting.add(
        'L2:L1000',
        FormulaRule(formula=['AND($L2="Yes",$N2="")'],
                    fill=PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type="solid"),
                    font=Font(color="FFFFFF", bold=True))
    )
    
    # Set column widths
    column_widths = {
        1: 12,  # RowID
        2: 12,  # Status
        3: 25,  # Source System
        4: 25,  # Destination System
        5: 30,  # PII Categories Transferred
        6: 18,  # Transfer Method
        7: 15,  # Transfer Frequency
        8: 30,  # Purpose of Transfer
        9: 25,  # Legal Basis
        10: 18, # Recipient Type
        11: 25, # Recipient Name
        12: 15, # Cross-Border Transfer?
        13: 20, # Destination Country
        14: 30, # Transfer Mechanism
        15: 15, # SCC Date
        16: 20, # TIA Completed?
        17: 35, # Security Measures
        18: 15, # Last Updated
        19: 20, # Updated By
        20: 35, # Notes
        21: 20, # Evidence Reference
        22: 25  # Review Comments
    }
    set_column_widths(ws, column_widths)
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Add sample row with formulas
    ws['A2'] = '=TEXT(ROW()-1,"FLOW-000")'
    ws['R2'] = '=TODAY()'
    
    # Protect sheet
    protect_sheet(ws, allow_edit_ranges=['B2:V1000'])
    
    return ws


def create_ropa_sheet(wb):
    """Create Sheet 4: Record of Processing Activities (ROPA) - GDPR Art. 30"""
    ws = wb.create_sheet("4. ROPA (Record of Processing)")
    
    headers = [
        "RowID",
        "Status",
        "Processing Activity Name",
        "Purpose of Processing",
        "Legal Basis (Art. 6)",
        "Legal Basis (Art. 9 if special category)",
        "Legal Basis Justification",
        "Categories of Data Subjects",
        "Categories of Personal Data",
        "Contains Special Category Data?",
        "Specific Special Category Types",
        "Categories of Recipients (Internal)",
        "Categories of Recipients (Processors)",
        "Categories of Recipients (Third Parties)",
        "Categories of Recipients (Public Authorities)",
        "Transfers to Third Countries?",
        "Third Countries",
        "Transfer Safeguards",
        "Retention Period",
        "Retention Justification",
        "Deletion Method",
        "Technical Security Measures",
        "Organizational Security Measures",
        "Data Sources",
        "Data Subject Rights Supported",
        "DPIA Completed?",
        "DPIA Reference",
        "LIA Completed? (if Leg. Interest)",
        "LIA Reference",
        "Last Updated",
        "Updated By",
        "Notes",
        "Evidence Reference",
        "Review Comments"
    ]
    
    create_header_row(ws, headers)
    
    # Add data validation
    add_data_validation(ws, "B2:B1000", STATUS_OPTIONS)
    add_data_validation(ws, "E2:E1000", LEGAL_BASIS_ART6)
    add_data_validation(ws, "F2:F1000", LEGAL_BASIS_ART9)
    add_data_validation(ws, "J2:J1000", ["Yes", "No"])
    add_data_validation(ws, "P2:P1000", ["Yes", "No"])
    add_data_validation(ws, "Z2:Z1000", ["Yes", "No", "N/A"])
    add_data_validation(ws, "AB2:AB1000", ["Yes", "No", "N/A"])
    
    # Add conditional formatting
    add_status_conditional_formatting(ws, 2, 2, 1000)
    
    # Highlight missing legal basis
    ws.conditional_formatting.add(
        'E2:E1000',
        FormulaRule(formula=['AND($B2<>"Not Started",$E2="")'],
                    fill=PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type="solid"),
                    font=Font(color="FFFFFF", bold=True))
    )
    
    # Highlight special category without Art. 9 basis
    ws.conditional_formatting.add(
        'F2:F1000',
        FormulaRule(formula=['AND($J2="Yes",$F2="N/A (No special category data)")'],
                    fill=PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type="solid"),
                    font=Font(color="FFFFFF", bold=True))
    )
    
    # Set column widths
    column_widths = {
        1: 12,  # RowID
        2: 12,  # Status
        3: 30,  # Processing Activity Name
        4: 40,  # Purpose of Processing
        5: 25,  # Legal Basis (Art. 6)
        6: 30,  # Legal Basis (Art. 9)
        7: 40,  # Legal Basis Justification
        8: 25,  # Categories of Data Subjects
        9: 40,  # Categories of Personal Data
        10: 18, # Contains Special Category Data?
        11: 35, # Specific Special Category Types
        12: 30, # Recipients (Internal)
        13: 35, # Recipients (Processors)
        14: 30, # Recipients (Third Parties)
        15: 30, # Recipients (Public Authorities)
        16: 15, # Transfers to Third Countries?
        17: 25, # Third Countries
        18: 35, # Transfer Safeguards
        19: 30, # Retention Period
        20: 35, # Retention Justification
        21: 25, # Deletion Method
        22: 40, # Technical Security Measures
        23: 40, # Organizational Security Measures
        24: 30, # Data Sources
        25: 35, # Data Subject Rights Supported
        26: 15, # DPIA Completed?
        27: 20, # DPIA Reference
        28: 18, # LIA Completed?
        29: 20, # LIA Reference
        30: 15, # Last Updated
        31: 20, # Updated By
        32: 40, # Notes
        33: 20, # Evidence Reference
        34: 30  # Review Comments
    }
    set_column_widths(ws, column_widths)
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Add sample row with formulas
    ws['A2'] = '=TEXT(ROW()-1,"ROPA-000")'
    ws['AD2'] = '=TODAY()'
    
    # Protect sheet
    protect_sheet(ws, allow_edit_ranges=['B2:AH1000'])
    
    return ws


def create_gaps_sheet(wb):
    """Create Sheet 5: PII Discovery Gaps"""
    ws = wb.create_sheet("5. PII Discovery Gaps")
    
    headers = [
        "GapID",
        "Status",
        "Gap Type",
        "Gap Description",
        "System/Activity Affected",
        "Risk Level",
        "Risk Justification",
        "Remediation Action",
        "Remediation Owner",
        "Target Completion Date",
        "Actual Completion Date",
        "Date Identified",
        "Identified By",
        "Last Updated",
        "Updated By",
        "Notes",
        "Evidence Reference",
        "Review Comments"
    ]
    
    create_header_row(ws, headers)
    
    # Add data validation
    add_data_validation(ws, "B2:B1000", ["Open", "In Progress", "Completed", "Accepted"])
    add_data_validation(ws, "C2:C1000", GAP_TYPES)
    add_data_validation(ws, "F2:F1000", RISK_LEVELS)
    
    # Add conditional formatting
    # Status formatting
    ws.conditional_formatting.add(
        'B2:B1000',
        CellIsRule(operator='equal', formula=['"Open"'],
                   fill=PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'B2:B1000',
        CellIsRule(operator='equal', formula=['"In Progress"'],
                   fill=PatternFill(start_color=COLOR_IN_PROGRESS, end_color=COLOR_IN_PROGRESS, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        'B2:B1000',
        CellIsRule(operator='equal', formula=['"Completed"'],
                   fill=PatternFill(start_color=COLOR_COMPLETE, end_color=COLOR_COMPLETE, fill_type="solid"))
    )
    
    # Risk level formatting
    add_risk_conditional_formatting(ws, 6, 2, 1000)
    
    # Overdue highlighting
    ws.conditional_formatting.add(
        'J2:J1000',
        FormulaRule(formula=['AND($J2<TODAY(),$B2<>"Completed")'],
                    fill=PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type="solid"),
                    font=Font(color="FFFFFF", bold=True))
    )
    
    # Set column widths
    column_widths = {
        1: 15,  # GapID
        2: 15,  # Status
        3: 25,  # Gap Type
        4: 45,  # Gap Description
        5: 30,  # System/Activity Affected
        6: 12,  # Risk Level
        7: 35,  # Risk Justification
        8: 45,  # Remediation Action
        9: 25,  # Remediation Owner
        10: 18, # Target Completion Date
        11: 18, # Actual Completion Date
        12: 15, # Date Identified
        13: 20, # Identified By
        14: 15, # Last Updated
        15: 20, # Updated By
        16: 40, # Notes
        17: 20, # Evidence Reference
        18: 30  # Review Comments
    }
    set_column_widths(ws, column_widths)
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Add sample row with formulas
    ws['A2'] = '=TEXT(YEAR(TODAY()),"0000")&"-GAP-"&TEXT(ROW()-1,"000")'
    ws['L2'] = '=TODAY()'
    ws['N2'] = '=TODAY()'
    
    # Protect sheet
    protect_sheet(ws, allow_edit_ranges=['B2:R1000'])
    
    return ws


def create_evidence_register_sheet(wb):
    """Create Sheet 6: Evidence Register"""
    ws = wb.create_sheet("6. Evidence Register")
    
    headers = [
        "EvidenceID",
        "Evidence Type",
        "Evidence Description",
        "Related System/Activity",
        "Related Sheet Reference",
        "Date Collected",
        "Collected By",
        "File Location/URL",
        "File Name",
        "Retention Period",
        "Confidentiality Level",
        "Last Updated",
        "Updated By",
        "Notes",
        "Review Comments"
    ]
    
    create_header_row(ws, headers)
    
    # Add data validation
    add_data_validation(ws, "B2:B1000", EVIDENCE_TYPES)
    add_data_validation(ws, "J2:J1000", ["Permanent", "3 years", "5 years", "7 years", "10 years"])
    add_data_validation(ws, "K2:K1000", ["Public", "Internal", "Confidential", "Restricted"])
    
    # Add conditional formatting for confidentiality
    ws.conditional_formatting.add(
        'K2:K1000',
        CellIsRule(operator='equal', formula=['"Restricted"'],
                   fill=PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type="solid"),
                   font=Font(color="FFFFFF", bold=True))
    )
    
    # Set column widths
    column_widths = {
        1: 18,  # EvidenceID
        2: 25,  # Evidence Type
        3: 40,  # Evidence Description
        4: 30,  # Related System/Activity
        5: 25,  # Related Sheet Reference
        6: 15,  # Date Collected
        7: 20,  # Collected By
        8: 40,  # File Location/URL
        9: 25,  # File Name
        10: 15, # Retention Period
        11: 18, # Confidentiality Level
        12: 15, # Last Updated
        13: 20, # Updated By
        14: 40, # Notes
        15: 30  # Review Comments
    }
    set_column_widths(ws, column_widths)
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Add sample row with formulas
    ws['A2'] = '="EVID-A534.1-"&TEXT(ROW()-1,"000")'
    ws['F2'] = '=TODAY()'
    ws['L2'] = '=TODAY()'
    
    # Protect sheet
    protect_sheet(ws, allow_edit_ranges=['B2:O1000'])
    
    return ws


def create_dashboard_sheet(wb):
    """Create Sheet 7: Dashboard (Compliance Metrics)"""
    ws = wb.create_sheet("7. Dashboard")
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "PII Identification & Classification Dashboard"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dashboard content (summary metrics)
    dashboard_rows = [
        ("", ""),
        ("Assessment Date:", f"=TEXT(TODAY(),\"DD.MM.YYYY\")"),
        ("", ""),
        ("SYSTEM INVENTORY METRICS:", ""),
        ("Total Systems Assessed:", "=COUNTA('2. PII System Inventory'!C2:C1000)"),
        ("Systems with Basic PII:", "=COUNTIF('2. PII System Inventory'!I2:I1000,\"Basic PII\")"),
        ("Systems with Sensitive PII:", "=COUNTIF('2. PII System Inventory'!I2:I1000,\"*Sensitive*\")"),
        ("Systems with Criminal Data:", "=COUNTIF('2. PII System Inventory'!I2:I1000,\"*Criminal*\")"),
        ("Systems Status - Not Started:", "=COUNTIF('2. PII System Inventory'!B2:B1000,\"Not Started\")"),
        ("Systems Status - In Progress:", "=COUNTIF('2. PII System Inventory'!B2:B1000,\"In Progress\")"),
        ("Systems Status - Complete:", "=COUNTIF('2. PII System Inventory'!B2:B1000,\"Complete\")"),
        ("Systems Status - Validated:", "=COUNTIF('2. PII System Inventory'!B2:B1000,\"Validated\")"),
        ("", ""),
        ("DATA FLOW METRICS:", ""),
        ("Total Data Flows Mapped:", "=COUNTA('3. PII Data Flow Mapping'!C2:C1000)"),
        ("Cross-Border Transfers:", "=COUNTIF('3. PII Data Flow Mapping'!L2:L1000,\"Yes\")"),
        ("Flows Missing Transfer Mechanism:", "=SUMPRODUCT(('3. PII Data Flow Mapping'!L2:L1000=\"Yes\")*('3. PII Data Flow Mapping'!N2:N1000=\"\"))"),
        ("", ""),
        ("ROPA METRICS:", ""),
        ("Total Processing Activities:", "=COUNTA('4. ROPA (Record of Processing)'!C2:C1000)"),
        ("Activities with Special Category Data:", "=COUNTIF('4. ROPA (Record of Processing)'!J2:J1000,\"Yes\")"),
        ("Activities Missing Legal Basis:", "=SUMPRODUCT(('4. ROPA (Record of Processing)'!B2:B1000<>\"Not Started\")*('4. ROPA (Record of Processing)'!E2:E1000=\"\"))"),
        ("Activities with Cross-Border Transfers:", "=COUNTIF('4. ROPA (Record of Processing)'!P2:P1000,\"Yes\")"),
        ("", ""),
        ("GAP ANALYSIS:", ""),
        ("Total Gaps Identified:", "=COUNTA('5. PII Discovery Gaps'!A2:A1000)"),
        ("Gaps - Critical Risk:", "=COUNTIF('5. PII Discovery Gaps'!F2:F1000,\"Critical\")"),
        ("Gaps - High Risk:", "=COUNTIF('5. PII Discovery Gaps'!F2:F1000,\"High\")"),
        ("Gaps - Medium Risk:", "=COUNTIF('5. PII Discovery Gaps'!F2:F1000,\"Medium\")"),
        ("Gaps - Low Risk:", "=COUNTIF('5. PII Discovery Gaps'!F2:F1000,\"Low\")"),
        ("Gaps - Open:", "=COUNTIF('5. PII Discovery Gaps'!B2:B1000,\"Open\")"),
        ("Gaps - In Progress:", "=COUNTIF('5. PII Discovery Gaps'!B2:B1000,\"In Progress\")"),
        ("Gaps - Completed:", "=COUNTIF('5. PII Discovery Gaps'!B2:B1000,\"Completed\")"),
        ("Gaps - Overdue:", "=SUMPRODUCT(('5. PII Discovery Gaps'!J2:J1000<TODAY())*('5. PII Discovery Gaps'!B2:B1000<>\"Completed\"))"),
        ("", ""),
        ("EVIDENCE METRICS:", ""),
        ("Total Evidence Items Collected:", "=COUNTA('6. Evidence Register'!A2:A1000)"),
        ("", ""),
        ("COMPLIANCE STATUS:", ""),
        ("Assessment Completion Rate:", "=IFERROR(TEXT(COUNTIF('2. PII System Inventory'!B2:B1000,\"Validated\")/COUNTA('2. PII System Inventory'!C2:C1000),\"0%\"),\"0%\")"),
        ("Critical/High Risk Gaps:", "=COUNTIF('5. PII Discovery Gaps'!F2:F1000,\"Critical\")+COUNTIF('5. PII Discovery Gaps'!F2:F1000,\"High\")"),
        ("", ""),
        ("NEXT ACTIONS:", ""),
        ("", "1. Review and remediate Critical/High risk gaps"),
        ("", "2. Validate all systems marked 'Complete' → 'Validated'"),
        ("", "3. Complete missing legal basis documentation"),
        ("", "4. Obtain Standard Contractual Clauses for cross-border transfers without safeguards"),
        ("", "5. Conduct DPIAs for high-risk processing activities"),
        ("", "6. Obtain executive approvals in Sheet 8 (Sign-Off)"),
    ]
    
    for row_num, (label, value) in enumerate(dashboard_rows, 3):
        ws.cell(row=row_num, column=1, value=label).font = Font(bold=True)
        if value:
            ws.cell(row=row_num, column=2, value=value)
    
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 60
    
    # Protect sheet (dashboard is read-only, formulas pull from other sheets)
    protect_sheet(ws)
    
    return ws


def create_signoff_sheet(wb):
    """Create Sheet 8: Approval & Sign-Off"""
    ws = wb.create_sheet("8. Approval & Sign-Off")
    
    # Title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "Assessment Approval & Sign-Off"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    headers = [
        "Signatory Role",
        "Signatory Name",
        "Signature / Electronic Approval",
        "Signature Date",
        "Approval Scope",
        "Comments",
        "Contact Email"
    ]
    
    create_header_row(ws, headers, start_row=3)
    
    # Add data validation
    add_data_validation(ws, "A4:A1000", SIGNATORY_ROLES)
    
    # Pre-populate required roles
    required_roles = [
        ("Assessment Lead (DPO / Privacy Officer)", "", "", "", "Assessment methodology, completeness, ROPA compliance", "", ""),
        ("Chief Information Security Officer (CISO)", "", "", "", "Technical accuracy, security measures documentation", "", ""),
        ("Legal / Compliance Officer", "", "", "", "Legal bases validity, regulatory compliance", "", ""),
        ("Executive Sponsor", "", "", "", "Final approval, gap remediation support", "", ""),
    ]
    
    for row_num, role_data in enumerate(required_roles, 4):
        for col_num, value in enumerate(role_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Set column widths
    column_widths = {
        1: 35,  # Signatory Role
        2: 25,  # Signatory Name
        3: 30,  # Signature
        4: 15,  # Signature Date
        5: 40,  # Approval Scope
        6: 40,  # Comments
        7: 30   # Contact Email
    }
    set_column_widths(ws, column_widths)
    
    # Conditional formatting - highlight unsigned
    ws.conditional_formatting.add(
        'D4:D1000',
        FormulaRule(formula=['$D4=""'],
                    fill=PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type="solid"))
    )
    
    # Protect sheet
    protect_sheet(ws, allow_edit_ranges=['A4:G1000'])
    
    return ws


# ============================================================================
# MAIN WORKBOOK GENERATION
# ============================================================================

def generate_workbook(output_path=None, date_suffix=None):
    """Generate complete assessment workbook"""
    
    # Determine output filename
    if date_suffix:
        date_str = date_suffix
    else:
        date_str = datetime.now().strftime("%Y%m%d")
    
    filename = f"ISMS-IMP-A.5.34.1_PII_Identification_Assessment_{date_str}.xlsx"
    
    if output_path:
        filepath = os.path.join(output_path, filename)
    else:
        filepath = filename
    
    logger.info(f"Generating PII Identification Assessment Workbook...")
    logger.info(f"Output: {filepath}")
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create all sheets
    logger.info("Creating Sheet 1: Instructions & Legend...")
    create_instructions_sheet(wb)
    
    logger.info("Creating Sheet 2: PII System Inventory...")
    create_system_inventory_sheet(wb)
    
    logger.info("Creating Sheet 3: PII Data Flow Mapping...")
    create_data_flow_mapping_sheet(wb)
    
    logger.info("Creating Sheet 4: ROPA (Record of Processing Activities)...")
    create_ropa_sheet(wb)
    
    logger.info("Creating Sheet 5: PII Discovery Gaps...")
    create_gaps_sheet(wb)
    
    logger.info("Creating Sheet 6: Evidence Register...")
    create_evidence_register_sheet(wb)
    
    logger.info("Creating Sheet 7: Dashboard...")
    create_dashboard_sheet(wb)
    
    logger.info("Creating Sheet 8: Approval & Sign-Off...")
    create_signoff_sheet(wb)
    
    # Save workbook
    logger.info("Saving workbook...")
    wb.save(filepath)
    
    logger.info(f"✓ Workbook generated successfully: {filepath}")
    logger.info(f"\nNext steps:")
    logger.info(f"1. Open the workbook and review instructions (Sheet 1)")
    logger.info(f"2. Begin PII system inventory (Sheet 2)")
    logger.info(f"3. Map data flows including cross-border transfers (Sheet 3)")
    logger.info(f"4. Create ROPA entries for all processing activities (Sheet 4)")
    logger.info(f"5. Document gaps and remediation plans (Sheet 5)")
    logger.info(f"6. Collect and register audit evidence (Sheet 6)")
    logger.info(f"7. Review dashboard metrics (Sheet 7)")
    logger.info(f"8. Obtain stakeholder approvals (Sheet 8)")
    logger.info(f"\nRefer to ISMS-IMP-A.5.34.1 Part 1 (User Guide) for detailed completion instructions.")
    
    return filepath


# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Generate ISMS-IMP-A.5.34.1 PII Identification Assessment Workbook",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage (generates in current directory with today's date)
  python3 generate_a5341_pii_identification_assessment.py
  
  # Specify output directory
  python3 generate_a5341_pii_identification_assessment.py --output /path/to/assessments
  
  # Specify custom date suffix
  python3 generate_a5341_pii_identification_assessment.py --date 20250128
  
  # Combine options
  python3 generate_a5341_pii_identification_assessment.py --output ./assessments --date 20250128

Output:
  ISMS_A_5_34_1_PII_Identification_Assessment_YYYYMMDD.xlsx

For detailed completion instructions, refer to:
  ISMS-IMP-A.5.34.1 Part 1: User Completion Guide
  ISMS-IMP-A.5.34.1 Part 2: Technical Specification
        """
    )
    
    parser.add_argument(
        '--output',
        type=str,
        help='Output directory path (default: current directory)'
    )
    
    parser.add_argument(
        '--date',
        type=str,
        help='Date suffix for filename in YYYYMMDD format (default: today)'
    )
    
    args = parser.parse_args()
    
    # Validate date format if provided
    if args.date:
        try:
            datetime.strptime(args.date, "%Y%m%d")
        except ValueError:
            logger.error("Error: Date must be in YYYYMMDD format (e.g., 20250128)")
            return 1
    
    # Validate output directory if provided
    if args.output and not os.path.exists(args.output):
        logger.error(f"Error: Output directory does not exist: {args.output}")
        return 1
    
    try:
        generate_workbook(output_path=args.output, date_suffix=args.date)
        return 0
    except Exception as e:
        logger.error(f"Error generating workbook: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
