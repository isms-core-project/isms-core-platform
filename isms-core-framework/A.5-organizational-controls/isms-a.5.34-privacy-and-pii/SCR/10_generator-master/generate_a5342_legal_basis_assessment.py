#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.34.2 - Legal Basis and Lawful Processing Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.34: Privacy and Protection of PII
Assessment Domain 2 of 7: Legal Basis for Processing, Lawful Processing Framework

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific data processing activities, legal frameworks, and
privacy assessment requirements.

Key customization areas:
1. Legal basis dropdown options (adapt to applicable jurisdictions)
2. Consent management processes (match your actual consent capture mechanisms)
3. Legitimate interest assessment criteria (align with organizational risk appetite)
4. Special category data definitions (verify against local regulations)
5. Stakeholder approval workflow (match organizational structure)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.34 Privacy and Protection of PII Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
and documenting the legal basis for all personal data processing activities.

**Purpose:**
Enables systematic documentation of lawful processing requirements under GDPR
Article 6, Swiss FADP Article 34, and related data protection regulations,
supporting evidence-based validation of legal basis for ISO 27001 Control A.5.34
compliance.

**Assessment Scope:**
- Legal basis inventory for all processing activities
- Legitimate Interest Assessments (LIA) with three-part balancing test
- Consent management and GDPR Article 7 compliance validation
- Special category data (sensitive PII) legal basis under GDPR Article 9
- Legal basis gap analysis and remediation tracking
- Evidence repository for audit readiness
- Data subject transparency and information requirements
- Cross-border transfer legal basis implications

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and legal framework reference
2. Legal Basis Inventory - Complete legal basis documentation for all processing
3. Legitimate Interest Assessments - LIA balancing tests (Purpose/Necessity/Balancing)
4. Consent Management - Consent validity assessment and withdrawal tracking
5. Legal Basis Gaps - Auto-populated gaps requiring remediation
6. Evidence Repository - Supporting documentation (consent logs, contracts, LIAs)
7. Dashboard - Compliance metrics and executive summary
8. Approval & Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with legal basis and consent mechanism dropdown lists
- Conditional formatting for compliance status and risk levels
- Automated gap detection for missing or invalid legal basis
- LIA balancing test with automated result calculation
- Consent validity checks against GDPR Article 7 requirements
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with ROPA from A.5.34.1 (PII Identification Assessment)

**Integration:**
This assessment:
- Requires ISMS-IMP-A.5.34.1 (PII Identification) as prerequisite
- Feeds into ISMS-IMP-A.5.34.3 (Data Subject Rights Management)
- Supports ISMS-IMP-A.5.34.5 (DPIA) legal basis requirements
- Consolidates into ISMS-IMP-A.5.34.7 (Privacy Compliance Dashboard)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl >= 3.0.0 (Python Excel library)
    - Python standard library (argparse, datetime, os)

Data Prerequisites:
    - ISMS-IMP-A.5.34.1 completed (PII Identification Assessment with ROPA)
    - Privacy notices and policies (for transparency verification)
    - Consent records (if consent-based processing exists)
    - Contracts (if contract-based processing exists)
    - Legal obligations register (if applicable)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a5342_legal_basis_assessment.py

Custom Output File:
    python3 generate_a5342_legal_basis_assessment.py \\
        --output /path/to/ISMS_A_5_34_2_Legal_Basis_Assessment_20241231.xlsx

Custom Date (for file naming):
    python3 generate_a5342_legal_basis_assessment.py \\
        --date 20241231

Full Example:
    python3 generate_a5342_legal_basis_assessment.py \\
        --output ./assessments/legal_basis_assessment.xlsx \\
        --date 20241231

Output:
    - Excel workbook with 8 worksheets
    - File naming convention: ISMS_A_5_34_2_Legal_Basis_Assessment_YYYYMMDD.xlsx
    - Size: Approximately 2-5 MB (scales with row limits)

--------------------------------------------------------------------------------
CUSTOMIZATION GUIDE
--------------------------------------------------------------------------------

Before first use, review and customize the following sections:

1. CONFIGURATION SECTION (lines ~200-400):
   - Legal basis dropdown values (add jurisdiction-specific options)
   - Consent mechanisms (match organizational processes)
   - Risk scoring criteria (align with risk appetite)
   - Color scheme (organizational branding)
   - Protection password (set organizational standard)

2. DROPDOWN LISTS (lines ~300-350):
   - GDPR Article 6 legal bases (add/modify as needed)
   - GDPR Article 9 special category bases (verify completeness)
   - Consent mechanisms (match actual capture methods)
   - Data subject information channels (add organizational channels)

3. VALIDATION RULES (embedded in sheet functions):
   - Legal basis selection logic
   - Consent validity checks
   - LIA balancing test calculations
   - Risk level scoring

4. APPROVAL WORKFLOW (Sheet 8 function):
   - Stakeholder roles (match organizational structure)
   - Approval routing (add/remove approvers)
   - Sign-off requirements (single/multi-signature)

--------------------------------------------------------------------------------
REGULATORY ALIGNMENT
--------------------------------------------------------------------------------

This assessment supports compliance with:

**European Union:**
- GDPR Articles 6 (Lawfulness of processing)
- GDPR Article 7 (Conditions for consent)
- GDPR Article 9 (Processing of special categories of personal data)
- GDPR Articles 13-14 (Information to be provided to data subject)
- GDPR Recital 47 (Legitimate interests)

**Switzerland:**
- Federal Act on Data Protection (FADP/nDSG) Article 6 (Principles)
- FADP Article 34 (Legal basis for federal bodies)
- Swiss Code of Obligations Article 328b (Employee personality protection)

**International Standards:**
- ISO/IEC 27001:2022 Control A.5.34 (Privacy and protection of PII)
- ISO/IEC 27701:2019 (Privacy Information Management System)
- ISO/IEC 29134 (Privacy impact assessment guidelines)

**Industry-Specific:**
- Financial services: Legal basis for AML/KYC processing
- Healthcare: HIPAA compliance for health data legal basis
- Telecommunications: ePrivacy Directive consent requirements

--------------------------------------------------------------------------------
SCRIPT METADATA
--------------------------------------------------------------------------------

Version:        1.0
Created:        2025-01-28
Last Modified:        [Date to be set]
Author:               [Organization] ISMS Implementation Team
Related Docs:   ISMS-IMP-A.5.34.2-Part1-UserGuide.md
                ISMS-IMP-A.5.34.2-Part2-TechSpec.md
                ISMS-POL-A.5.34 (Privacy and Protection of PII Policy)

Script Purpose: Generate Excel assessment workbook for legal basis documentation
Workbook Type:  Data entry + automated calculations + dashboard
Target Users:   Data Protection Officers, Privacy Officers, Legal Counsel,
                Business Process Owners, Compliance Teams

Quality:        Production-ready template requiring organizational customization
Testing:        Structural validation complete; functional testing required with
                organization-specific data

--------------------------------------------------------------------------------
CHANGE LOG
--------------------------------------------------------------------------------

Version 1.0 (2025-01-28):
    - Initial release
    - 8 worksheets covering complete legal basis assessment lifecycle
    - GDPR and Swiss FADP alignment
    - Legitimate Interest Assessment (LIA) framework
    - Consent management with GDPR Art. 7 validation
    - Gap analysis with risk-based prioritization
    - Evidence repository with cross-referencing
    - Dashboard with compliance metrics
    - Approval workflow with multi-stakeholder sign-off

Future Enhancements (Planned):
    - Multi-language support (German, French, Italian)
    - Enhanced LIA methodology (CNIL/ICO guidance integration)
    - Automated consent log import
    - Integration with DPA reporting requirements
    - Historical tracking for legal basis changes

--------------------------------------------------------------------------------
SUPPORT AND FEEDBACK
--------------------------------------------------------------------------------

For questions, issues, or enhancement requests:

Technical Issues:
    - Check Python and openpyxl versions meet requirements
    - Verify file write permissions for output directory
    - Review error messages for missing configuration

Compliance Questions:
    - Consult Data Protection Officer (DPO)
    - Review ISMS-POL-A.5.34 (Privacy Policy)
    - Reference Part 1 User Guide (Section 4: Common Questions)

Customization Support:
    - Review Part 2 Technical Specification
    - Examine "# CUSTOMIZE:" comment blocks
    - Test changes with sample data before production use

Regulatory Updates:
    - Monitor EDPB guidelines and supervisory authority guidance
    - Update dropdown values when regulations change
    - Revise LIA criteria based on case law and enforcement actions

================================================================================
"""

# ============================================================================
# IMPORTS
# ============================================================================

import argparse
import os
from datetime import datetime

# Excel generation library
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# ============================================================================
# CONFIGURATION - CUSTOMIZE THIS SECTION
# ============================================================================

# CUSTOMIZE: Update these values to match your organization's requirements

# Color Scheme (Organization-agnostic defaults)
# Adjust these hex codes to match your organization's brand guidelines
COLOR_HEADER = "003366"          # Dark blue header
COLOR_INSTRUCTION = "D6DCE4"     # Light blue instructions
COLOR_INPUT = "FFFFFF"           # White input cells
COLOR_CALCULATED = "F2F2F2"      # Light gray calculated cells
COLOR_COMPLIANT = "C6EFCE"       # Green - compliant status
COLOR_REVIEW = "FFEB9C"          # Yellow - requires review
COLOR_GAP = "FFC7CE"             # Red - gap/non-compliant
COLOR_CRITICAL = "C00000"        # Dark red - critical risk
COLOR_HIGH = "FF6600"            # Orange - high risk
COLOR_MEDIUM = "FFD966"          # Yellow - medium risk
COLOR_LOW = "9BC2E6"             # Light blue - low risk

# Sheet Protection Password
# CUSTOMIZE: Set a strong password for your organization
PROTECTION_PASSWORD = "privacy2024"

# File Naming
# CUSTOMIZE: Adjust naming convention if needed
FILE_PREFIX = "ISMS_A_5_34_2_Legal_Basis_Assessment"

# Dropdown List Values
# CUSTOMIZE: Add or modify legal bases based on applicable jurisdictions

# GDPR Article 6 Legal Bases
LEGAL_BASIS_ART6 = [
    "Consent (GDPR Art. 6(1)(a))",
    "Contract (GDPR Art. 6(1)(b))",
    "Legal Obligation (GDPR Art. 6(1)(c))",
    "Vital Interests (GDPR Art. 6(1)(d))",
    "Public Task (GDPR Art. 6(1)(e))",
    "Legitimate Interest (GDPR Art. 6(1)(f))"
]

# GDPR Article 9 Special Category Legal Bases
LEGAL_BASIS_ART9 = [
    "Explicit Consent (Art. 9(2)(a))",
    "Employment Law (Art. 9(2)(b))",
    "Vital Interests (Art. 9(2)(c))",
    "Legitimate Activities (Art. 9(2)(d))",
    "Public Data (Art. 9(2)(e))",
    "Legal Claims (Art. 9(2)(f))",
    "Substantial Public Interest (Art. 9(2)(g))",
    "Healthcare (Art. 9(2)(h))",
    "Public Health (Art. 9(2)(i))",
    "Archiving/Research (Art. 9(2)(j))"
]

# Data Subject Information Channels
DS_INFORMATION_CHANNELS = [
    "Privacy Notice (Website)",
    "Privacy Notice (Contract)",
    "Consent Form",
    "Employee Handbook",
    "Terms of Service",
    "Not Communicated"
]

# Consent Mechanisms
CONSENT_MECHANISMS = [
    "Website Form (Opt-In Checkbox)",
    "Double Opt-In (Email Confirmation)",
    "Written Signature",
    "Verbal Consent (Recorded)",
    "Implied Consent",
    "Not Documented"
]

# Consent Status Options
CONSENT_STATUS = [
    "Valid",
    "Invalid - Not Freely Given",
    "Invalid - Not Specific",
    "Invalid - Not Informed",
    "Invalid - Ambiguous",
    "Not Captured"
]

# Compliance Status Options
COMPLIANCE_STATUS = [
    "Compliant",
    "Requires Review",
    "Gap - No Legal Basis",
    "Gap - Invalid Consent",
    "Gap - Missing LIA"
]

# LIA - Beneficiary Options
LIA_BENEFICIARY = [
    "Controller",
    "Data Subject",
    "Third Party",
    "Multiple Parties"
]

# LIA - Assessment Options
LIA_LEGITIMACY = ["Legitimate", "Questionable", "Not Legitimate"]
LIA_NECESSITY = ["Necessary", "Questionable", "Not Necessary"]
LIA_EXPECTATIONS = ["Reasonable Expectation", "Unexpected", "Surprising"]
LIA_IMPACT = ["Minimal Impact", "Limited Impact", "Significant Impact", "High Impact"]
LIA_RESULT = [
    "Pass - Legitimate Interest Prevails",
    "Pass with Conditions",
    "Fail - Data Subject's Rights Prevail"
]

# Risk Level Options
RISK_LEVELS = ["Critical", "High", "Medium", "Low"]

# Remediation Status
REMEDIATION_STATUS = ["Not Started", "In Progress", "Complete", "Blocked"]

# Evidence Types
EVIDENCE_TYPES = [
    "Consent Records",
    "Customer Contract",
    "Employment Contract",
    "Processor Agreement (DPA)",
    "Legal Obligation (Statute)",
    "Legitimate Interest Assessment (LIA)",
    "Privacy Notice",
    "Terms of Service",
    "Employee Handbook",
    "Consent Form Template",
    "Consent Log Export",
    "Other"
]

# Verification Status
VERIFICATION_STATUS = ["Valid", "Expired", "Requires Update", "Under Review"]

# Approval Status
APPROVAL_STATUS = ["Approved", "Approved with Conditions", "Requires Revision", "Rejected"]

# Yes/No Options
YES_NO = ["Yes", "No"]
YES_NO_UNCERTAIN = ["Yes", "No", "Uncertain"]
YES_NO_PENDING = ["Yes", "No", "Pending"]

# Default Row Limits (adjust for scalability)
# CUSTOMIZE: Increase for large organizations
MAX_ROWS_INVENTORY = 10000       # Sheet 2: Legal Basis Inventory
MAX_ROWS_LIA = 1000              # Sheet 3: LIA
MAX_ROWS_CONSENT = 1000          # Sheet 4: Consent Management
MAX_ROWS_GAPS = 1000             # Sheet 5: Gaps
MAX_ROWS_EVIDENCE = 1000         # Sheet 6: Evidence

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def create_header_row(ws, headers, start_col=1, start_row=1, header_color=COLOR_HEADER):
    """
    Create formatted header row with consistent styling.
    
    Args:
        ws: Worksheet object
        headers: List of header text strings
        start_col: Starting column (1-indexed)
        start_row: Starting row (1-indexed)
        header_color: Hex color code for header background
    """
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border_side = Side(style='thin', color='000000')
    header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for idx, header_text in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + idx)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        cell.protection = Protection(locked=True)


def set_column_widths(ws, column_widths):
    """
    Set column widths based on dictionary mapping.
    
    Args:
        ws: Worksheet object
        column_widths: Dict mapping column letters to width values
                       Example: {'A': 15, 'B': 40, 'C': 30}
    """
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width


def add_dropdown_validation(ws, col_range, values, allow_blank=False, error_title="Invalid Input", error_message="Please select from dropdown list"):
    """
    Add data validation dropdown to specified range.
    
    Args:
        ws: Worksheet object
        col_range: Range string (e.g., "D2:D10000")
        values: List of dropdown values
        allow_blank: Whether to allow blank cells
        error_title: Title for error alert
        error_message: Message for error alert
    """
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(values)}"',
        allow_blank=allow_blank,
        showErrorMessage=True,
        errorTitle=error_title,
        error=error_message
    )
    dv.add(col_range)
    ws.add_data_validation(dv)


def add_conditional_formatting_status(ws, col_range, status_type="compliance"):
    """
    Add conditional formatting for status columns.
    
    Args:
        ws: Worksheet object
        col_range: Range string (e.g., "N2:N10000")
        status_type: Type of status formatting ('compliance', 'risk', 'approval')
    """
    if status_type == "compliance":
        # Green for Compliant
        ws.conditional_formatting.add(
            col_range,
            CellIsRule(operator='equal', formula=['"Compliant"'], fill=PatternFill(start_color=COLOR_COMPLIANT, end_color=COLOR_COMPLIANT, fill_type='solid'))
        )
        # Yellow for Requires Review
        ws.conditional_formatting.add(
            col_range,
            CellIsRule(operator='equal', formula=['"Requires Review"'], fill=PatternFill(start_color=COLOR_REVIEW, end_color=COLOR_REVIEW, fill_type='solid'))
        )
        # Red for any Gap
        ws.conditional_formatting.add(
            col_range,
            CellIsRule(operator='beginsWith', formula=['"Gap"'], fill=PatternFill(start_color=COLOR_GAP, end_color=COLOR_GAP, fill_type='solid'))
        )
    
    elif status_type == "risk":
        # Critical
        ws.conditional_formatting.add(
            col_range,
            CellIsRule(operator='equal', formula=['"Critical"'], fill=PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type='solid'), font=Font(color='FFFFFF', bold=True))
        )
        # High
        ws.conditional_formatting.add(
            col_range,
            CellIsRule(operator='equal', formula=['"High"'], fill=PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type='solid'), font=Font(color='FFFFFF'))
        )
        # Medium
        ws.conditional_formatting.add(
            col_range,
            CellIsRule(operator='equal', formula=['"Medium"'], fill=PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type='solid'))
        )
        # Low
        ws.conditional_formatting.add(
            col_range,
            CellIsRule(operator='equal', formula=['"Low"'], fill=PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type='solid'))
        )
    
    elif status_type == "approval":
        # Approved
        ws.conditional_formatting.add(
            col_range,
            CellIsRule(operator='equal', formula=['"Approved"'], fill=PatternFill(start_color=COLOR_COMPLIANT, end_color=COLOR_COMPLIANT, fill_type='solid'))
        )
        # Approved with Conditions
        ws.conditional_formatting.add(
            col_range,
            CellIsRule(operator='equal', formula=['"Approved with Conditions"'], fill=PatternFill(start_color=COLOR_REVIEW, end_color=COLOR_REVIEW, fill_type='solid'))
        )
        # Requires Revision or Rejected
        ws.conditional_formatting.add(
            col_range,
            CellIsRule(operator='equal', formula=['"Requires Revision"'], fill=PatternFill(start_color=COLOR_GAP, end_color=COLOR_GAP, fill_type='solid'))
        )
        ws.conditional_formatting.add(
            col_range,
            CellIsRule(operator='equal', formula=['"Rejected"'], fill=PatternFill(start_color=COLOR_GAP, end_color=COLOR_GAP, fill_type='solid'))
        )


def protect_sheet(ws, password=PROTECTION_PASSWORD, allow_select_locked=True, allow_select_unlocked=True, allow_filter=True, allow_sort=True):
    """
    Apply sheet protection with specified permissions.
    
    Args:
        ws: Worksheet object
        password: Protection password
        allow_select_locked: Allow selecting locked cells
        allow_select_unlocked: Allow selecting unlocked cells
        allow_filter: Allow using autofilter
        allow_sort: Allow sorting
    """
    ws.protection.sheet = True
    ws.protection.password = password
    ws.protection.selectLockedCells = allow_select_locked
    ws.protection.selectUnlockedCells = allow_select_unlocked
    ws.protection.autoFilter = allow_filter
    ws.protection.sort = allow_sort


# ============================================================================
# SHEET CREATION FUNCTIONS
# ============================================================================

def create_sheet1_instructions(wb):
    """
    Create Sheet 1: Instructions & Legend
    
    Embedded user guide and assessment methodology reference.
    """
    ws = wb.active
    ws.title = "1. Instructions & Legend"
    
    # Title
    ws['A1'] = "ISMS-IMP-A.5.34.2 - Legal Basis and Lawful Processing Assessment"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "Assessment Instructions and Legal Basis Selection Framework"
    ws['A2'].font = Font(name='Calibri', size=11, italic=True)
    ws.merge_cells('A2:F2')
    
    # Section 1: Assessment Overview
    row = 4
    ws[f'A{row}'] = "1. ASSESSMENT OVERVIEW"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='003366')
    
    row += 2
    overview_text = """This assessment documents the legal basis for ALL personal data processing activities to ensure compliance with:
• GDPR Article 6 (Lawfulness of processing)
• Swiss FADP Article 34 (Legal basis)
• ISO/IEC 27001:2022 Control A.5.34 (Privacy and protection of PII)

CRITICAL: Processing without documented legal basis is a fundamental data protection violation."""
    
    ws[f'A{row}'] = overview_text
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(f'A{row}:F{row+3}')
    ws.row_dimensions[row].height = 80
    
    # Section 2: Legal Basis Selection Framework
    row += 5
    ws[f'A{row}'] = "2. LEGAL BASIS SELECTION FRAMEWORK (GDPR Article 6)"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='003366')
    
    row += 2
    framework_text = """Legal basis is NOT a free choice. Selection must align with the ACTUAL purpose and nature of processing.

CONSENT (Art. 6(1)(a)): 
• Data subject has genuine choice (can refuse without detriment)
• Freely given, specific, informed, unambiguous
• Examples: Marketing to prospects, optional profiling, research participation

CONTRACT (Art. 6(1)(b)):
• Processing is objectively NECESSARY to perform contract obligations
• Without processing, contract cannot be fulfilled
• Examples: Shipping address for delivery, payment details for transactions
• NOT valid for "related to contract" - must be strictly necessary

LEGAL OBLIGATION (Art. 6(1)(c)):
• Specific law REQUIRES processing (cite statute)
• Examples: Tax records, employee social security, AML checks
• NOT valid for "best practices" - must be legal mandate

VITAL INTERESTS (Art. 6(1)(d)):
• Life-threatening emergency only
• Data subject cannot provide consent (unconscious, incapacitated)
• Rarely used

PUBLIC TASK (Art. 6(1)(e)):
• Public authority performing official function
• NOT applicable to private sector

LEGITIMATE INTEREST (Art. 6(1)(f)):
• Processing benefits controller or third party
• Balancing test passed (LIA completed - see Sheet 3)
• Data subject's interests DO NOT override controller's interests
• Examples: Fraud prevention, network security, direct marketing (with opt-out)
• Requires documented LIA"""
    
    ws[f'A{row}'] = framework_text
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(f'A{row}:F{row+23}')
    ws.row_dimensions[row].height = 380
    
    # Section 3: Special Category Data (GDPR Art. 9)
    row += 25
    ws[f'A{row}'] = "3. SPECIAL CATEGORY DATA (GDPR Article 9)"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='003366')
    
    row += 2
    special_text = """Special category PII requires ADDITIONAL legal basis under GDPR Art. 9:
• Racial or ethnic origin
• Political opinions
• Religious or philosophical beliefs
• Trade union membership
• Genetic data
• Biometric data (for unique identification)
• Health data
• Sex life or sexual orientation

IMPORTANT: General legal basis (Art. 6) + Special category basis (Art. 9) BOTH required."""
    
    ws[f'A{row}'] = special_text
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(f'A{row}:F{row+10}')
    ws.row_dimensions[row].height = 180
    
    # Section 4: Dropdown Reference
    row += 12
    ws[f'A{row}'] = "4. DROPDOWN OPTIONS REFERENCE"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='003366')
    
    row += 2
    ws[f'A{row}'] = "Legal Basis (GDPR Art. 6):"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    for basis in LEGAL_BASIS_ART6:
        ws[f'A{row}'] = f"• {basis}"
        row += 1
    
    row += 1
    ws[f'A{row}'] = "Special Category Legal Basis (GDPR Art. 9):"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    for basis in LEGAL_BASIS_ART9:
        ws[f'A{row}'] = f"• {basis}"
        row += 1
    
    # Set column widths
    set_column_widths(ws, {'A': 15, 'B': 20, 'C': 20, 'D': 20, 'E': 20, 'F': 20})
    
    # Lock all cells (read-only instructions)
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.protection = Protection(locked=True)
    
    # Protect sheet
    protect_sheet(ws)


def create_sheet2_legal_basis_inventory(wb):
    """
    Create Sheet 2: Legal Basis Inventory
    
    Complete legal basis documentation for all processing activities.
    """
    ws = wb.create_sheet("2. Legal Basis Inventory")
    
    # Headers
    headers = [
        "Activity ID",
        "Processing Activity",
        "Processing Purpose",
        "Legal Basis (GDPR Art. 6)",
        "Legal Basis Justification",
        "Data Subject Information",
        "Special Category Data?",
        "GDPR Art. 9 Legal Basis",
        "Consent Status",
        "Consent Mechanism",
        "LIA Required?",
        "LIA Reference",
        "Legal Basis Evidence",
        "Compliance Status",
        "Remediation Plan"
    ]
    
    create_header_row(ws, headers)
    
    # Set column widths
    column_widths = {
        'A': 15,  # Activity ID
        'B': 40,  # Processing Activity
        'C': 50,  # Processing Purpose
        'D': 30,  # Legal Basis
        'E': 60,  # Justification
        'F': 30,  # DS Information
        'G': 15,  # Special Category
        'H': 30,  # Art. 9 Basis
        'I': 20,  # Consent Status
        'J': 30,  # Consent Mechanism
        'K': 15,  # LIA Required
        'L': 20,  # LIA Reference
        'M': 50,  # Evidence
        'N': 20,  # Compliance Status
        'O': 60   # Remediation Plan
    }
    set_column_widths(ws, column_widths)
    
    # Add data validation (dropdowns)
    add_dropdown_validation(ws, f'D2:D{MAX_ROWS_INVENTORY}', LEGAL_BASIS_ART6, allow_blank=False, 
                           error_title="Invalid Legal Basis", error_message="Please select legal basis from dropdown list")
    
    add_dropdown_validation(ws, f'F2:F{MAX_ROWS_INVENTORY}', DS_INFORMATION_CHANNELS, allow_blank=False,
                           error_title="Invalid Channel", error_message="Please select data subject information channel")
    
    add_dropdown_validation(ws, f'G2:G{MAX_ROWS_INVENTORY}', YES_NO, allow_blank=False,
                           error_title="Invalid Input", error_message="Select Yes or No")
    
    add_dropdown_validation(ws, f'H2:H{MAX_ROWS_INVENTORY}', LEGAL_BASIS_ART9, allow_blank=True,
                           error_title="Invalid Art. 9 Basis", error_message="Please select Art. 9 legal basis if special category data = Yes")
    
    add_dropdown_validation(ws, f'I2:I{MAX_ROWS_INVENTORY}', CONSENT_STATUS, allow_blank=True,
                           error_title="Invalid Consent Status", error_message="Select consent status from dropdown")
    
    add_dropdown_validation(ws, f'J2:J{MAX_ROWS_INVENTORY}', CONSENT_MECHANISMS, allow_blank=True,
                           error_title="Invalid Mechanism", error_message="Select consent mechanism from dropdown")
    
    # Add formulas for calculated columns
    # Column N: Compliance Status (auto-calculated based on legal basis completeness)
    for row in range(2, MAX_ROWS_INVENTORY + 1):
        # Complex nested IF formula to determine compliance status
        compliance_formula = (
            f'=IF(D{row}="", "Gap - No Legal Basis", '
            f'IF(AND(D{row}="Consent (GDPR Art. 6(1)(a))", I{row}="Not Captured"), "Gap - Invalid Consent", '
            f'IF(AND(D{row}="Legitimate Interest (GDPR Art. 6(1)(f))", L{row}=""), "Gap - Missing LIA", '
            f'IF(AND(G{row}="Yes", H{row}=""), "Gap - Missing Art. 9 Basis", '
            f'IF(OR(I{row}="Invalid - Not Freely Given", I{row}="Invalid - Not Specific", I{row}="Invalid - Not Informed", I{row}="Invalid - Ambiguous"), "Gap - Invalid Consent", '
            f'"Compliant")))))'
        )
        ws[f'N{row}'] = compliance_formula
        ws[f'N{row}'].protection = Protection(locked=True)
    
    # Column K: LIA Required (auto-filled if legitimate interest)
    for row in range(2, MAX_ROWS_INVENTORY + 1):
        lia_formula = f'=IF(D{row}="Legitimate Interest (GDPR Art. 6(1)(f))", "Yes / Required (Sheet 3)", "")'
        ws[f'K{row}'] = lia_formula
        ws[f'K{row}'].protection = Protection(locked=True)
    
    # Unlock input cells (A-M, O)
    for row in range(2, MAX_ROWS_INVENTORY + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'M', 'O']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
    
    # Add conditional formatting
    add_conditional_formatting_status(ws, f'N2:N{MAX_ROWS_INVENTORY}', status_type="compliance")
    
    # Highlight special category data
    ws.conditional_formatting.add(
        f'G2:G{MAX_ROWS_INVENTORY}',
        CellIsRule(operator='equal', formula=['"Yes"'], fill=PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type='solid'), font=Font(bold=True))
    )
    
    # Highlight missing legal basis
    ws.conditional_formatting.add(
        f'D2:D{MAX_ROWS_INVENTORY}',
        FormulaRule(formula=[f'D2=""'], fill=PatternFill(start_color=COLOR_GAP, end_color=COLOR_GAP, fill_type='solid'))
    )
    
    # Highlight invalid consent
    ws.conditional_formatting.add(
        f'I2:I{MAX_ROWS_INVENTORY}',
        CellIsRule(operator='containsText', formula=['"Invalid"'], fill=PatternFill(start_color=COLOR_GAP, end_color=COLOR_GAP, fill_type='solid'))
    )
    
    # Protect sheet
    protect_sheet(ws)


def create_sheet3_lia(wb):
    """
    Create Sheet 3: Legitimate Interest Assessments (LIA)
    
    Three-part balancing test for GDPR Art. 6(1)(f) processing.
    """
    ws = wb.create_sheet("3. Legitimate Interest Assessments")
    
    # Headers
    headers = [
        "LIA ID",
        "Activity ID\n(from Sheet 2)",
        "Legitimate Interest",
        "Beneficiary",
        "Legitimacy Assessment",
        "Necessity Assessment",
        "Alternative Methods Considered",
        "Necessity Conclusion",
        "Data Subject's Interests",
        "Data Subject Expectations",
        "Impact Assessment",
        "Safeguards Implemented",
        "Balancing Test Result",
        "Assessor Name",
        "Assessment Date",
        "DPO Approval",
        "Approval Date"
    ]
    
    create_header_row(ws, headers)
    
    # Set column widths
    column_widths = {
        'A': 15,  # LIA ID
        'B': 15,  # Activity ID
        'C': 50,  # Legitimate Interest
        'D': 20,  # Beneficiary
        'E': 20,  # Legitimacy
        'F': 50,  # Necessity Assessment
        'G': 50,  # Alternatives
        'H': 20,  # Necessity Conclusion
        'I': 50,  # DS Interests
        'J': 20,  # DS Expectations
        'K': 20,  # Impact
        'L': 50,  # Safeguards
        'M': 30,  # Result
        'N': 25,  # Assessor
        'O': 15,  # Assessment Date
        'P': 15,  # DPO Approval
        'Q': 15   # Approval Date
    }
    set_column_widths(ws, column_widths)
    
    # Add data validation (dropdowns)
    add_dropdown_validation(ws, f'D2:D{MAX_ROWS_LIA}', LIA_BENEFICIARY, allow_blank=False)
    add_dropdown_validation(ws, f'E2:E{MAX_ROWS_LIA}', LIA_LEGITIMACY, allow_blank=False)
    add_dropdown_validation(ws, f'H2:H{MAX_ROWS_LIA}', LIA_NECESSITY, allow_blank=False)
    add_dropdown_validation(ws, f'J2:J{MAX_ROWS_LIA}', LIA_EXPECTATIONS, allow_blank=False)
    add_dropdown_validation(ws, f'K2:K{MAX_ROWS_LIA}', LIA_IMPACT, allow_blank=False)
    add_dropdown_validation(ws, f'P2:P{MAX_ROWS_LIA}', YES_NO_PENDING, allow_blank=False)
    
    # Add formulas for Balancing Test Result (Column M)
    for row in range(2, MAX_ROWS_LIA + 1):
        # Auto-calculate balancing test result based on legitimacy, necessity, and impact
        balancing_formula = (
            f'=IF(AND(E{row}="Legitimate", H{row}="Necessary", K{row}="Minimal Impact"), "Pass - Legitimate Interest Prevails", '
            f'IF(AND(E{row}="Legitimate", H{row}="Necessary", OR(K{row}="Limited Impact", K{row}="Significant Impact"), L{row}<>""), "Pass with Conditions", '
            f'IF(OR(E{row}="Not Legitimate", H{row}="Not Necessary", K{row}="High Impact"), "Fail - Data Subject\'\'s Rights Prevail", '
            f'"Pass with Conditions")))'
        )
        ws[f'M{row}'] = balancing_formula
        ws[f'M{row}'].protection = Protection(locked=True)
    
    # Unlock input cells (all except M)
    for row in range(2, MAX_ROWS_LIA + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'N', 'O', 'P', 'Q']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
    
    # Add conditional formatting
    # Balancing Test Result
    ws.conditional_formatting.add(
        f'M2:M{MAX_ROWS_LIA}',
        CellIsRule(operator='equal', formula=['"Pass - Legitimate Interest Prevails"'], fill=PatternFill(start_color=COLOR_COMPLIANT, end_color=COLOR_COMPLIANT, fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'M2:M{MAX_ROWS_LIA}',
        CellIsRule(operator='equal', formula=['"Pass with Conditions"'], fill=PatternFill(start_color=COLOR_REVIEW, end_color=COLOR_REVIEW, fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'M2:M{MAX_ROWS_LIA}',
        CellIsRule(operator='equal', formula=['"Fail - Data Subject\'s Rights Prevail"'], fill=PatternFill(start_color=COLOR_GAP, end_color=COLOR_GAP, fill_type='solid'))
    )
    
    # High impact processing
    ws.conditional_formatting.add(
        f'K2:K{MAX_ROWS_LIA}',
        CellIsRule(operator='equal', formula=['"High Impact"'], fill=PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type='solid'), font=Font(bold=True))
    )
    ws.conditional_formatting.add(
        f'K2:K{MAX_ROWS_LIA}',
        CellIsRule(operator='equal', formula=['"Significant Impact"'], fill=PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type='solid'), font=Font(bold=True))
    )
    
    # DPO Approval status
    add_conditional_formatting_status(ws, f'P2:P{MAX_ROWS_LIA}', status_type="approval")
    
    # Protect sheet
    protect_sheet(ws)


def create_sheet4_consent_management(wb):
    """
    Create Sheet 4: Consent Management
    
    Consent validity assessment and withdrawal tracking.
    """
    ws = wb.create_sheet("4. Consent Management")
    
    # Headers
    headers = [
        "Consent ID",
        "Activity ID\n(from Sheet 2)",
        "Consent Purpose",
        "Consent Date",
        "Consent Mechanism",
        "Freely Given?",
        "Specific?",
        "Informed?",
        "Unambiguous?",
        "Documented?",
        "Consent Validity",
        "Withdrawal Mechanism Available?",
        "Withdrawal Mechanism Description",
        "Withdrawal Processing Time",
        "Consent Records Location",
        "Audit Trail Available?",
        "Evidence Reference"
    ]
    
    create_header_row(ws, headers)
    
    # Set column widths
    column_widths = {
        'A': 15, 'B': 15, 'C': 40, 'D': 15, 'E': 30,
        'F': 15, 'G': 15, 'H': 15, 'I': 15, 'J': 15,
        'K': 25, 'L': 15, 'M': 50, 'N': 20, 'O': 40,
        'P': 15, 'Q': 20
    }
    set_column_widths(ws, column_widths)
    
    # Add data validation (dropdowns)
    add_dropdown_validation(ws, f'E2:E{MAX_ROWS_CONSENT}', CONSENT_MECHANISMS, allow_blank=False)
    add_dropdown_validation(ws, f'F2:J{MAX_ROWS_CONSENT}', YES_NO_UNCERTAIN, allow_blank=False)
    add_dropdown_validation(ws, f'L2:L{MAX_ROWS_CONSENT}', YES_NO, allow_blank=False)
    add_dropdown_validation(ws, f'P2:P{MAX_ROWS_CONSENT}', YES_NO, allow_blank=False)
    
    # Add formulas for Consent Validity (Column K)
    for row in range(2, MAX_ROWS_CONSENT + 1):
        # Auto-calculate consent validity based on GDPR Art. 7 requirements
        validity_formula = (
            f'=IF(F{row}="No", "Invalid - Not Freely Given", '
            f'IF(G{row}="No", "Invalid - Not Specific", '
            f'IF(H{row}="No", "Invalid - Not Informed", '
            f'IF(I{row}="No", "Invalid - Ambiguous", '
            f'IF(J{row}="No", "Invalid - Not Documented", '
            f'IF(AND(F{row}="Yes", G{row}="Yes", H{row}="Yes", I{row}="Yes", J{row}="Yes"), "Valid", '
            f'"Requires Review"))))))'
        )
        ws[f'K{row}'] = validity_formula
        ws[f'K{row}'].protection = Protection(locked=True)
    
    # Unlock input cells (all except K)
    for row in range(2, MAX_ROWS_CONSENT + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'M', 'N', 'O', 'P', 'Q']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
    
    # Add conditional formatting
    # Consent Validity
    ws.conditional_formatting.add(
        f'K2:K{MAX_ROWS_CONSENT}',
        CellIsRule(operator='equal', formula=['"Valid"'], fill=PatternFill(start_color=COLOR_COMPLIANT, end_color=COLOR_COMPLIANT, fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'K2:K{MAX_ROWS_CONSENT}',
        CellIsRule(operator='containsText', formula=['"Invalid"'], fill=PatternFill(start_color=COLOR_GAP, end_color=COLOR_GAP, fill_type='solid'))
    )
    
    # GDPR Art. 7 compliance (F-J)
    for col in ['F', 'G', 'H', 'I', 'J']:
        ws.conditional_formatting.add(
            f'{col}2:{col}{MAX_ROWS_CONSENT}',
            CellIsRule(operator='equal', formula=['"No"'], fill=PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type='solid'))
        )
    
    # Withdrawal mechanism missing
    ws.conditional_formatting.add(
        f'L2:L{MAX_ROWS_CONSENT}',
        CellIsRule(operator='equal', formula=['"No"'], fill=PatternFill(start_color=COLOR_GAP, end_color=COLOR_GAP, fill_type='solid'), font=Font(bold=True))
    )
    
    # Protect sheet
    protect_sheet(ws)


def create_sheet5_gaps(wb):
    """
    Create Sheet 5: Legal Basis Gaps
    
    Auto-populated gaps from Sheet 2 requiring remediation.
    """
    ws = wb.create_sheet("5. Legal Basis Gaps")
    
    # Headers
    headers = [
        "Activity ID",
        "Processing Activity",
        "Current Legal Basis",
        "Gap Type",
        "Risk Level",
        "Remediation Priority",
        "Remediation Action",
        "Remediation Owner",
        "Target Date",
        "Status",
        "Completion Date",
        "Verification Notes"
    ]
    
    create_header_row(ws, headers)
    
    # Set column widths
    column_widths = {
        'A': 15, 'B': 40, 'C': 30, 'D': 30, 'E': 15, 'F': 15,
        'G': 60, 'H': 25, 'I': 15, 'J': 15, 'K': 15, 'L': 50
    }
    set_column_widths(ws, column_widths)
    
    # Note: This sheet would ideally be auto-populated from Sheet 2 using formulas or Python
    # For template, provide structure and let users manually populate or use consolidation script
    
    # Add data validation (dropdowns)
    add_dropdown_validation(ws, f'E2:E{MAX_ROWS_GAPS}', RISK_LEVELS, allow_blank=False)
    add_dropdown_validation(ws, f'F2:F{MAX_ROWS_GAPS}', RISK_LEVELS, allow_blank=False)
    add_dropdown_validation(ws, f'J2:J{MAX_ROWS_GAPS}', REMEDIATION_STATUS, allow_blank=False)
    
    # Lock columns A-D (would be auto-populated from Sheet 2)
    for row in range(2, MAX_ROWS_GAPS + 1):
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].protection = Protection(locked=True)
    
    # Unlock remediation planning columns (F-L)
    for row in range(2, MAX_ROWS_GAPS + 1):
        for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
    
    # Add conditional formatting
    add_conditional_formatting_status(ws, f'E2:E{MAX_ROWS_GAPS}', status_type="risk")
    add_conditional_formatting_status(ws, f'F2:F{MAX_ROWS_GAPS}', status_type="risk")
    
    # Remediation status
    ws.conditional_formatting.add(
        f'J2:J{MAX_ROWS_GAPS}',
        CellIsRule(operator='equal', formula=['"Complete"'], fill=PatternFill(start_color=COLOR_COMPLIANT, end_color=COLOR_COMPLIANT, fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'J2:J{MAX_ROWS_GAPS}',
        CellIsRule(operator='equal', formula=['"In Progress"'], fill=PatternFill(start_color=COLOR_REVIEW, end_color=COLOR_REVIEW, fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'J2:J{MAX_ROWS_GAPS}',
        CellIsRule(operator='equal', formula=['"Blocked"'], fill=PatternFill(start_color=COLOR_GAP, end_color=COLOR_GAP, fill_type='solid'))
    )
    
    # Overdue remediation
    ws.conditional_formatting.add(
        f'I2:I{MAX_ROWS_GAPS}',
        FormulaRule(formula=[f'AND(I2<TODAY(), J2<>"Complete")'], fill=PatternFill(start_color=COLOR_GAP, end_color=COLOR_GAP, fill_type='solid'), font=Font(bold=True))
    )
    
    # Protect sheet
    protect_sheet(ws)


def create_sheet6_evidence(wb):
    """
    Create Sheet 6: Evidence Repository
    
    Supporting documentation for legal basis claims.
    """
    ws = wb.create_sheet("6. Evidence Repository")
    
    # Headers
    headers = [
        "Evidence ID",
        "Evidence Type",
        "Description",
        "File Location",
        "Linked Activity IDs",
        "Verified By",
        "Verification Date",
        "Verification Status",
        "Notes"
    ]
    
    create_header_row(ws, headers)
    
    # Set column widths
    column_widths = {
        'A': 15, 'B': 30, 'C': 50, 'D': 60, 'E': 30,
        'F': 25, 'G': 15, 'H': 20, 'I': 50
    }
    set_column_widths(ws, column_widths)
    
    # Add data validation (dropdowns)
    add_dropdown_validation(ws, f'B2:B{MAX_ROWS_EVIDENCE}', EVIDENCE_TYPES, allow_blank=False)
    add_dropdown_validation(ws, f'H2:H{MAX_ROWS_EVIDENCE}', VERIFICATION_STATUS, allow_blank=False)
    
    # Unlock all input cells
    for row in range(2, MAX_ROWS_EVIDENCE + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
    
    # Add conditional formatting
    ws.conditional_formatting.add(
        f'H2:H{MAX_ROWS_EVIDENCE}',
        CellIsRule(operator='equal', formula=['"Valid"'], fill=PatternFill(start_color=COLOR_COMPLIANT, end_color=COLOR_COMPLIANT, fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'H2:H{MAX_ROWS_EVIDENCE}',
        CellIsRule(operator='equal', formula=['"Expired"'], fill=PatternFill(start_color=COLOR_GAP, end_color=COLOR_GAP, fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'H2:H{MAX_ROWS_EVIDENCE}',
        CellIsRule(operator='equal', formula=['"Requires Update"'], fill=PatternFill(start_color=COLOR_GAP, end_color=COLOR_GAP, fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'H2:H{MAX_ROWS_EVIDENCE}',
        CellIsRule(operator='equal', formula=['"Under Review"'], fill=PatternFill(start_color=COLOR_REVIEW, end_color=COLOR_REVIEW, fill_type='solid'))
    )
    
    # Highlight missing verification
    ws.conditional_formatting.add(
        f'F2:G{MAX_ROWS_EVIDENCE}',
        FormulaRule(formula=[f'F2=""'], fill=PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type='solid'))
    )
    
    # Protect sheet
    protect_sheet(ws)


def create_sheet7_dashboard(wb):
    """
    Create Sheet 7: Dashboard
    
    Compliance metrics and executive summary with auto-calculated KPIs.
    """
    ws = wb.create_sheet("7. Dashboard")
    
    # Title
    ws['A1'] = "Legal Basis Compliance Dashboard"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='003366')
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True)
    ws.merge_cells('A2:F2')
    
    # Section 1: Legal Basis Coverage
    row = 4
    ws[f'A{row}'] = "LEGAL BASIS COVERAGE"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='003366')
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    metrics = [
        ("Total Processing Activities", f"=COUNTA('2. Legal Basis Inventory'!A2:A{MAX_ROWS_INVENTORY})"),
        ("Activities with Legal Basis", f"=COUNTIF('2. Legal Basis Inventory'!D2:D{MAX_ROWS_INVENTORY}, \"<>\")"),
        ("Coverage Percentage", f"=IF(B5=0, 0, B6/B5)"),
        ("Gap Count", f"=COUNTIF('2. Legal Basis Inventory'!N2:N{MAX_ROWS_INVENTORY}, \"Gap*\")")
    ]
    
    for metric_name, formula in metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = formula
        ws[f'B{row}'].protection = Protection(locked=True)
        if "Percentage" in metric_name:
            ws[f'B{row}'].number_format = '0.0%'
        row += 1
    
    # Section 2: Legal Basis Distribution
    row += 2
    ws[f'A{row}'] = "LEGAL BASIS DISTRIBUTION"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='003366')
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    for basis in LEGAL_BASIS_ART6:
        ws[f'A{row}'] = basis.split('(')[0].strip()  # Extract name without article reference
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"=COUNTIF('2. Legal Basis Inventory'!D2:D{MAX_ROWS_INVENTORY}, \"{basis}*\")"
        ws[f'B{row}'].protection = Protection(locked=True)
        row += 1
    
    # Section 3: Consent Management Metrics
    row += 2
    ws[f'A{row}'] = "CONSENT MANAGEMENT"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='003366')
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    consent_metrics = [
        ("Total Consents", f"=COUNTA('4. Consent Management'!A2:A{MAX_ROWS_CONSENT})"),
        ("Valid Consents", f"=COUNTIF('4. Consent Management'!K2:K{MAX_ROWS_CONSENT}, \"Valid\")"),
        ("Invalid Consents", f"=COUNTIFS('4. Consent Management'!K2:K{MAX_ROWS_CONSENT}, \"Invalid*\")"),
        ("Consent Validity Rate", f"=IF(B{row}=0, 0, B{row+1}/B{row})"),
        ("Consents Without Withdrawal", f"=COUNTIF('4. Consent Management'!L2:L{MAX_ROWS_CONSENT}, \"No\")")
    ]
    
    for metric_name, formula in consent_metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = formula
        ws[f'B{row}'].protection = Protection(locked=True)
        if "Rate" in metric_name:
            ws[f'B{row}'].number_format = '0.0%'
        row += 1
    
    # Section 4: Legitimate Interest Assessments
    row += 2
    ws[f'A{row}'] = "LEGITIMATE INTEREST ASSESSMENTS (LIA)"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='003366')
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    lia_metrics = [
        ("Total LIAs Required", f"=COUNTIF('2. Legal Basis Inventory'!K2:K{MAX_ROWS_INVENTORY}, \"Yes*\")"),
        ("LIAs Completed", f"=COUNTA('3. Legitimate Interest Assessments'!A2:A{MAX_ROWS_LIA})"),
        ("LIA Completion Rate", f"=IF(B{row}=0, 0, B{row+1}/B{row})"),
        ("LIAs Passed", f"=COUNTIF('3. Legitimate Interest Assessments'!M2:M{MAX_ROWS_LIA}, \"Pass*\")"),
        ("LIAs Failed", f"=COUNTIF('3. Legitimate Interest Assessments'!M2:M{MAX_ROWS_LIA}, \"Fail*\")"),
        ("LIAs Pending DPO Approval", f"=COUNTIF('3. Legitimate Interest Assessments'!P2:P{MAX_ROWS_LIA}, \"Pending\")")
    ]
    
    for metric_name, formula in lia_metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = formula
        ws[f'B{row}'].protection = Protection(locked=True)
        if "Rate" in metric_name:
            ws[f'B{row}'].number_format = '0.0%'
        row += 1
    
    # Section 5: Gap Analysis
    row += 2
    ws[f'A{row}'] = "GAP ANALYSIS"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='003366')
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    gap_metrics = [
        ("Critical Gaps", f"=COUNTIF('5. Legal Basis Gaps'!E2:E{MAX_ROWS_GAPS}, \"Critical\")"),
        ("High Priority Gaps", f"=COUNTIF('5. Legal Basis Gaps'!E2:E{MAX_ROWS_GAPS}, \"High\")"),
        ("Medium Priority Gaps", f"=COUNTIF('5. Legal Basis Gaps'!E2:E{MAX_ROWS_GAPS}, \"Medium\")"),
        ("Low Priority Gaps", f"=COUNTIF('5. Legal Basis Gaps'!E2:E{MAX_ROWS_GAPS}, \"Low\")"),
        ("Gaps Remediated", f"=COUNTIF('5. Legal Basis Gaps'!J2:J{MAX_ROWS_GAPS}, \"Complete\")"),
        ("Remediation Progress", f"=IF(SUM(B{row}:B{row+3})=0, 0, B{row+4}/SUM(B{row}:B{row+3}))")
    ]
    
    for metric_name, formula in gap_metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = formula
        ws[f'B{row}'].protection = Protection(locked=True)
        if "Progress" in metric_name:
            ws[f'B{row}'].number_format = '0.0%'
        row += 1
    
    # Section 6: Overall Compliance Score
    row += 2
    ws[f'A{row}'] = "OVERALL COMPLIANCE SCORE"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='003366')
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    # Calculate average of key metrics
    ws[f'A{row}'] = "Compliance Score"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'] = "=AVERAGE(B7, B27, B36)"  # Coverage %, Consent Validity Rate, LIA Completion Rate
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].protection = Protection(locked=True)
    
    # Add conditional formatting for compliance score
    ws.conditional_formatting.add(
        f'B{row}',
        CellIsRule(operator='greaterThanOrEqual', formula=['0.9'], fill=PatternFill(start_color=COLOR_COMPLIANT, end_color=COLOR_COMPLIANT, fill_type='solid'), font=Font(bold=True, color='FFFFFF'))
    )
    ws.conditional_formatting.add(
        f'B{row}',
        CellIsRule(operator='between', formula=['0.7', '0.9'], fill=PatternFill(start_color=COLOR_REVIEW, end_color=COLOR_REVIEW, fill_type='solid'), font=Font(bold=True))
    )
    ws.conditional_formatting.add(
        f'B{row}',
        CellIsRule(operator='between', formula=['0.5', '0.7'], fill=PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type='solid'), font=Font(bold=True))
    )
    ws.conditional_formatting.add(
        f'B{row}',
        CellIsRule(operator='lessThan', formula=['0.5'], fill=PatternFill(start_color=COLOR_GAP, end_color=COLOR_GAP, fill_type='solid'), font=Font(bold=True, color='FFFFFF'))
    )
    
    # Set column widths
    set_column_widths(ws, {'A': 40, 'B': 20, 'C': 20, 'D': 20, 'E': 20, 'F': 20})
    
    # Lock all cells (dashboard is fully calculated)
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.protection = Protection(locked=True)
    
    # Protect sheet
    protect_sheet(ws)


def create_sheet8_approval(wb):
    """
    Create Sheet 8: Approval & Sign-Off
    
    Stakeholder review and approval workflow.
    """
    ws = wb.create_sheet("8. Approval & Sign-Off")
    
    # Title
    ws['A1'] = "Legal Basis Assessment - Approval and Sign-Off"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:E1')
    
    # Section 1: Assessment Completion
    row = 4
    ws[f'A{row}'] = "ASSESSMENT COMPLETION"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='003366')
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = "Assessor Name:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].protection = Protection(locked=False)
    
    row += 1
    ws[f'A{row}'] = "Role:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].protection = Protection(locked=False)
    
    row += 1
    ws[f'A{row}'] = "Completion Date:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].protection = Protection(locked=False)
    ws[f'B{row}'].number_format = 'YYYY-MM-DD'
    
    # Section 2: Stakeholder Reviews
    row += 3
    ws[f'A{row}'] = "STAKEHOLDER REVIEWS"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='003366')
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:E{row}')
    
    # Stakeholders
    stakeholders = [
        "Data Protection Officer (DPO)",
        "Legal Counsel",
        "Business Process Owner",
        "CISO / Chief Privacy Officer"
    ]
    
    row += 2
    # Header row for stakeholder table
    ws[f'A{row}'] = "Stakeholder"
    ws[f'B{row}'] = "Name"
    ws[f'C{row}'] = "Date"
    ws[f'D{row}'] = "Approval Status"
    ws[f'E{row}'] = "Comments"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
    
    row += 1
    start_approval_row = row
    for stakeholder in stakeholders:
        ws[f'A{row}'] = stakeholder
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].protection = Protection(locked=False)
        ws[f'C{row}'].protection = Protection(locked=False)
        ws[f'C{row}'].number_format = 'YYYY-MM-DD'
        ws[f'D{row}'].protection = Protection(locked=False)
        ws[f'E{row}'].protection = Protection(locked=False)
        row += 1
    
    # Add dropdown for approval status
    add_dropdown_validation(ws, f'D{start_approval_row}:D{row-1}', APPROVAL_STATUS, allow_blank=False)
    
    # Add conditional formatting for approval status
    add_conditional_formatting_status(ws, f'D{start_approval_row}:D{row-1}', status_type="approval")
    
    # Section 3: Final Approval
    row += 2
    ws[f'A{row}'] = "FINAL APPROVAL"
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='003366')
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_INSTRUCTION, end_color=COLOR_INSTRUCTION, fill_type='solid')
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = "Final Approver:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].protection = Protection(locked=False)
    
    row += 1
    ws[f'A{row}'] = "Title:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].protection = Protection(locked=False)
    
    row += 1
    ws[f'A{row}'] = "Approval Date:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].protection = Protection(locked=False)
    ws[f'B{row}'].number_format = 'YYYY-MM-DD'
    
    row += 1
    approval_date_row = row - 1
    ws[f'A{row}'] = "Next Review Date:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"=DATE(YEAR(B{approval_date_row})+1, MONTH(B{approval_date_row}), DAY(B{approval_date_row}))"
    ws[f'B{row}'].number_format = 'YYYY-MM-DD'
    ws[f'B{row}'].protection = Protection(locked=True)
    
    row += 1
    ws[f'A{row}'] = "Signature:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].protection = Protection(locked=False)
    
    # Set column widths
    set_column_widths(ws, {'A': 30, 'B': 25, 'C': 15, 'D': 25, 'E': 50})
    
    # Protect sheet (lock all except input fields - already set unlocked above)
    protect_sheet(ws)


# ============================================================================
# MAIN WORKBOOK GENERATION FUNCTION
# ============================================================================

def generate_legal_basis_assessment_workbook(output_file, date_str=None):
    """
    Generate complete Legal Basis Assessment Excel workbook.
    
    Args:
        output_file: Path to output Excel file
        date_str: Optional date string for file naming (YYYYMMDD format)
    
    Returns:
        Path to generated workbook
    """
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.5.34.2 - Legal Basis Assessment Workbook Generator")
    logger.info("=" * 80)
    logger.info("")
    
    # Create workbook
    logger.info("Creating workbook...")
    wb = Workbook()
    
    # Generate sheets
    logger.info("Generating Sheet 1: Instructions & Legend...")
    create_sheet1_instructions(wb)
    
    logger.info("Generating Sheet 2: Legal Basis Inventory...")
    create_sheet2_legal_basis_inventory(wb)
    
    logger.info("Generating Sheet 3: Legitimate Interest Assessments...")
    create_sheet3_lia(wb)
    
    logger.info("Generating Sheet 4: Consent Management...")
    create_sheet4_consent_management(wb)
    
    logger.info("Generating Sheet 5: Legal Basis Gaps...")
    create_sheet5_gaps(wb)
    
    logger.info("Generating Sheet 6: Evidence Repository...")
    create_sheet6_evidence(wb)
    
    logger.info("Generating Sheet 7: Dashboard...")
    create_sheet7_dashboard(wb)
    
    logger.info("Generating Sheet 8: Approval & Sign-Off...")
    create_sheet8_approval(wb)
    
    # Save workbook
    logger.info("")
    logger.info(f"Saving workbook to: {output_file}")
    wb.save(output_file)
    
    logger.info("")
    logger.info("=" * 80)
    logger.info("Workbook generation complete!")
    logger.info("=" * 80)
    logger.info("")
    logger.info("Generated workbook structure:")
    logger.info("  1. Instructions & Legend")
    logger.info("  2. Legal Basis Inventory")
    logger.info("  3. Legitimate Interest Assessments")
    logger.info("  4. Consent Management")
    logger.info("  5. Legal Basis Gaps")
    logger.info("  6. Evidence Repository")
    logger.info("  7. Dashboard")
    logger.info("  8. Approval & Sign-Off")
    logger.info("")
    logger.info("Next steps:")
    logger.info("  1. Open workbook in Excel or LibreOffice")
    logger.info("  2. Review Sheet 1 (Instructions) for assessment guidance")
    logger.info("  3. Complete Sheet 2 (Legal Basis Inventory) for all processing activities")
    logger.info("  4. Complete Sheet 3 (LIA) for legitimate interest processing")
    logger.info("  5. Complete Sheet 4 (Consent Management) for consent-based processing")
    logger.info("  6. Review Sheet 7 (Dashboard) for compliance metrics")
    logger.info("  7. Obtain stakeholder approvals in Sheet 8")
    logger.info("")
    logger.info(f"Output file: {output_file}")
    logger.info("")
    
    return output_file


# ============================================================================
# COMMAND LINE INTERFACE
# ============================================================================

def main():
    """Command-line interface for workbook generation."""
    
    parser = argparse.ArgumentParser(
        description='Generate ISMS-IMP-A.5.34.2 Legal Basis Assessment Excel Workbook',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate with default filename
  python3 generate_a5342_legal_basis_assessment.py
  
  # Generate with custom output path
  python3 generate_a5342_legal_basis_assessment.py --output ./assessments/legal_basis.xlsx
  
  # Generate with custom date
  python3 generate_a5342_legal_basis_assessment.py --date 20241231
  
  # Full custom
  python3 generate_a5342_legal_basis_assessment.py --output ./legal_basis.xlsx --date 20241231

File naming convention:
  ISMS_A_5_34_2_Legal_Basis_Assessment_YYYYMMDD.xlsx

For detailed instructions, see:
  - ISMS-IMP-A.5.34.2-Part1-UserGuide.md
  - ISMS-IMP-A.5.34.2-Part2-TechSpec.md
        """
    )
    
    parser.add_argument(
        '--output',
        type=str,
        default=None,
        help='Output file path (default: ISMS_A_5_34_2_Legal_Basis_Assessment_YYYYMMDD.xlsx)'
    )
    
    parser.add_argument(
        '--date',
        type=str,
        default=None,
        help='Date string for filename in YYYYMMDD format (default: today\'s date)'
    )
    
    args = parser.parse_args()
    
    # Determine date string
    if args.date:
        date_str = args.date
    else:
        date_str = datetime.now().strftime('%Y%m%d')
    
    # Determine output filename
    if args.output:
        output_file = args.output
    else:
        output_file = f"{FILE_PREFIX}_{date_str}.xlsx"
    
    # Ensure output directory exists
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Generate workbook
    try:
        generated_file = generate_legal_basis_assessment_workbook(output_file, date_str)
        logger.info(f"SUCCESS: Workbook generated successfully")
        logger.info(f"Location: {os.path.abspath(generated_file)}")
        return 0
    except Exception as e:
        logger.error(f"ERROR: Failed to generate workbook")
        logger.error(f"Error details: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1



# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.34.2"
WORKBOOK_NAME = "Legal Basis and Lawful Processing Assessment"
CONTROL_ID = "A.5.34"
CONTROL_NAME = "Privacy and Protection of PII"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


if __name__ == '__main__':
    exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
