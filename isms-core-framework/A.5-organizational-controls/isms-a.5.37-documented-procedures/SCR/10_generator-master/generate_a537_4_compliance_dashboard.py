#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.37.4 - Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.37: Documented Operating Procedures
Assessment Domain 4 of 4: Compliance Dashboard and Executive Reporting

Reference Pattern: Based on ISMS-IMP-A.5.37.4 specification
================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.37.4"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.5.37"
CONTROL_NAME = "Documented Operating Procedures"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "kpi_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "metric_good": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "metric_warning": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "metric_bad": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "border": border_thin,
    }
    return styles


def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Executive_Dashboard",
        "Category_Analysis",
        "Review_Status",
        "Quality_Overview",
        "Gap_Tracking",
        "Change_Activity",
        "Trend_History",
        "Data_Sources",
        "Instructions",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def create_executive_dashboard_sheet(ws, styles):
    """Create the Executive_Dashboard sheet - KPIs and summary metrics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "A.5.37 DOCUMENTED PROCEDURES - EXECUTIVE DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # KPI Section Header
    ws.merge_cells("A3:G3")
    ws["A3"] = "KEY PERFORMANCE INDICATORS"
    ws["A3"].font = styles["kpi_header"]["font"]
    ws["A3"].fill = styles["kpi_header"]["fill"]
    ws["A3"].alignment = styles["kpi_header"]["alignment"]

    kpi_headers = ["Metric", "Current Value", "Target", "Status", "Trend"]
    for col, header in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    kpis = [
        ("Total Procedures Documented", "", "-", "", ""),
        ("Procedures Approved (%)", "", ">95%", "", ""),
        ("Reviews Current (%)", "", ">95%", "", ""),
        ("Average Quality Score", "", "≥4.0", "", ""),
        ("Open Gaps", "", "0", "", ""),
        ("Overdue Reviews", "", "0", "", ""),
    ]

    row = 6
    for metric, value, target, status, trend in kpis:
        ws.cell(row=row, column=1, value=metric).border = styles["border"]
        for c in range(2, 6):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        ws.cell(row=row, column=3, value=target)
        row += 1

    # Compliance Status Section
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "COMPLIANCE STATUS BY CATEGORY"
    ws[f"A{row}"].font = styles["kpi_header"]["font"]
    ws[f"A{row}"].fill = styles["kpi_header"]["fill"]
    ws[f"A{row}"].alignment = styles["kpi_header"]["alignment"]

    row += 2
    cat_headers = ["Category", "Total", "Approved", "Current", "Compliance %", "Status"]
    for col, header in enumerate(cat_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    categories = [
        "System Operations",
        "Security Operations",
        "Facility Operations",
        "Change Management",
        "Recovery Operations",
        "User Management",
    ]

    row += 1
    for cat in categories:
        ws.cell(row=row, column=1, value=cat).border = styles["border"]
        for c in range(2, 7):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).border = styles["border"]
    for c in range(2, 7):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.border = styles["border"]

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15


def create_category_analysis_sheet(ws, styles):
    """Create the Category_Analysis sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "CATEGORY ANALYSIS - Breakdown by Procedure Category"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Category", 25),
        ("Total_Count", 14),
        ("Documented", 14),
        ("Approved", 14),
        ("Current", 14),
        ("Compliance_%", 14),
        ("Avg_Quality", 14),
        ("Open_Gaps", 14),
        ("Status", 14),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    categories = [
        "System Operations",
        "Security Operations",
        "Facility Operations",
        "Change Management",
        "Recovery Operations",
        "User Management",
        "Other",
    ]

    row = 4
    for cat in categories:
        ws.cell(row=row, column=1, value=cat).border = styles["border"]
        for c in range(2, 10):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).border = styles["border"]
    for c in range(2, 10):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.border = styles["border"]

    ws.freeze_panes = "B4"


def create_review_status_sheet(ws, styles):
    """Create the Review_Status sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "REVIEW STATUS - Review Compliance Overview"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Status", 18),
        ("Count", 14),
        ("Percentage", 14),
        ("Critical_Count", 16),
        ("High_Count", 14),
        ("Medium_Count", 16),
        ("Low_Count", 14),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    statuses = ["OVERDUE", "DUE SOON", "CURRENT"]

    row = 4
    for status in statuses:
        ws.cell(row=row, column=1, value=status).border = styles["border"]
        for c in range(2, 8):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Top 10 Overdue Section
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "TOP 10 OVERDUE PROCEDURES"
    ws[f"A{row}"].font = styles["kpi_header"]["font"]
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].alignment = styles["kpi_header"]["alignment"]

    row += 2
    overdue_headers = ["Rank", "Procedure_ID", "Procedure_Name", "Days_Overdue", "Owner", "Escalation_Level"]
    for col, header in enumerate(overdue_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    row += 1
    for i in range(1, 11):
        ws.cell(row=row, column=1, value=i).border = styles["border"]
        for c in range(2, 7):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1


def create_quality_overview_sheet(ws, styles):
    """Create the Quality_Overview sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "QUALITY OVERVIEW - Assessment Summary"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Rating", 18),
        ("Count", 14),
        ("Percentage", 14),
        ("Avg_Clarity", 14),
        ("Avg_Completeness", 16),
        ("Avg_Accuracy", 14),
        ("Avg_Usability", 14),
        ("Avg_Maintainability", 18),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ratings = ["Excellent", "Good", "Adequate", "Needs Improvement", "Poor"]

    row = 4
    for rating in ratings:
        ws.cell(row=row, column=1, value=rating).border = styles["border"]
        for c in range(2, 9):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1


def create_gap_tracking_sheet(ws, styles):
    """Create the Gap_Tracking sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "GAP TRACKING - Remediation Status"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Severity", 14),
        ("Open", 12),
        ("In_Progress", 14),
        ("Closed", 12),
        ("Total", 12),
        ("Closure_Rate", 14),
        ("Avg_Days_Open", 14),
        ("Overdue", 12),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    severities = ["Critical", "High", "Medium", "Low"]

    row = 4
    for severity in severities:
        ws.cell(row=row, column=1, value=severity).border = styles["border"]
        for c in range(2, 9):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).border = styles["border"]
    for c in range(2, 9):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.border = styles["border"]


def create_change_activity_sheet(ws, styles):
    """Create the Change_Activity sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "CHANGE ACTIVITY - Change Request Metrics"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Period", 16),
        ("Submitted", 14),
        ("Approved", 14),
        ("Rejected", 14),
        ("Implemented", 14),
        ("Approval_Rate", 14),
        ("Avg_Cycle_Time", 16),
        ("Emergency_Count", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(4, 16):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]


def create_trend_history_sheet(ws, styles):
    """Create the Trend_History sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "TREND HISTORY - Historical Performance"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Period", 16),
        ("Total_Procedures", 16),
        ("Compliance_%", 14),
        ("Avg_Quality", 14),
        ("Open_Gaps", 12),
        ("Overdue_Reviews", 16),
        ("CR_Volume", 14),
        ("Notes", 35),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(4, 24):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]


def create_data_sources_sheet(ws, styles):
    """Create the Data_Sources sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "DATA SOURCES - Source Workbook Links"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Source_Workbook", 45),
        ("Sheet", 25),
        ("Data_Range", 20),
        ("Last_Updated", 16),
        ("Refresh_Method", 16),
        ("Status", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    sources = [
        ("ISMS-IMP-A.5.37.1_Procedure_Inventory_Assessment.xlsx", "Procedure_Inventory", "A4:P103"),
        ("ISMS-IMP-A.5.37.2_Procedure_Quality_Assessment.xlsx", "Quality_Assessment", "A4:N103"),
        ("ISMS-IMP-A.5.37.3_Procedure_Review_and_Update_Tracking.xlsx", "Review_Schedule", "A4:N103"),
        ("ISMS-IMP-A.5.37.3_Procedure_Review_and_Update_Tracking.xlsx", "Change_Requests", "A4:N103"),
    ]

    dv_method = DataValidation(
        type="list",
        formula1='"Manual,Automated"',
        allow_blank=False
    )
    ws.add_data_validation(dv_method)

    dv_status = DataValidation(
        type="list",
        formula1='"Connected,Error,Pending"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    row = 4
    for workbook, sheet, range_ref in sources:
        ws.cell(row=row, column=1, value=workbook).border = styles["border"]
        ws.cell(row=row, column=2, value=sheet).border = styles["border"]
        ws.cell(row=row, column=3, value=range_ref).border = styles["border"]
        for c in range(4, 7):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_method.add(ws.cell(row=row, column=5))
        dv_status.add(ws.cell(row=row, column=6))
        row += 1


def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Dashboard Purpose"
    ws["A3"].font = Font(bold=True, size=12)

    ws["A4"] = (
        "This dashboard consolidates procedure management metrics from A.5.37.1-3 workbooks "
        "into executive-ready visualisations. It provides real-time visibility into procedure "
        "documentation status, quality scores, review compliance, and gap remediation progress."
    )
    ws["A4"].alignment = Alignment(wrap_text=True)

    ws["A6"] = "Data Refresh Process"
    ws["A6"].font = Font(bold=True, size=12)

    refresh_steps = [
        "1. Open all source workbooks (A.5.37.1, A.5.37.2, A.5.37.3)",
        "2. Update external data links in this dashboard",
        "3. Refresh all pivot tables and formulas",
        "4. Verify calculation accuracy",
        "5. Update Last_Refreshed timestamp in Data_Sources sheet",
    ]

    row = 7
    for step in refresh_steps:
        ws[f"A{row}"] = step
        row += 1

    ws[f"A{row+1}"] = "Target Metrics"
    ws[f"A{row+1}"].font = Font(bold=True, size=12)

    targets = [
        ("Overall Compliance", ">95%"),
        ("Review Currency", ">95%"),
        ("Average Quality Score", "≥4.0"),
        ("Open Gaps", "0"),
        ("Overdue Reviews", "0"),
    ]

    row += 2
    for metric, target in targets:
        ws[f"A{row}"] = metric
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = target
        row += 1

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 20


def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/9] Creating Executive_Dashboard sheet...")
        create_executive_dashboard_sheet(wb["Executive_Dashboard"], styles)

        logger.info("[2/9] Creating Category_Analysis sheet...")
        create_category_analysis_sheet(wb["Category_Analysis"], styles)

        logger.info("[3/9] Creating Review_Status sheet...")
        create_review_status_sheet(wb["Review_Status"], styles)

        logger.info("[4/9] Creating Quality_Overview sheet...")
        create_quality_overview_sheet(wb["Quality_Overview"], styles)

        logger.info("[5/9] Creating Gap_Tracking sheet...")
        create_gap_tracking_sheet(wb["Gap_Tracking"], styles)

        logger.info("[6/9] Creating Change_Activity sheet...")
        create_change_activity_sheet(wb["Change_Activity"], styles)

        logger.info("[7/9] Creating Trend_History sheet...")
        create_trend_history_sheet(wb["Trend_History"], styles)

        logger.info("[8/9] Creating Data_Sources sheet...")
        create_data_sources_sheet(wb["Data_Sources"], styles)

        logger.info("[9/9] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.37.4 specification
# =============================================================================
