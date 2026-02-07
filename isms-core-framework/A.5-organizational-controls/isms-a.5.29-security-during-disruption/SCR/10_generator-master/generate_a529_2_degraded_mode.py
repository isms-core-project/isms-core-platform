#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.29.2 - Degraded Mode Security Requirements Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.29: Information Security During Disruption
Assessment Domain 2 of 4: Degraded Mode Security Requirements

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for documenting
degraded mode security requirements, break-glass procedures, and security debt.

**Purpose:**
Enables systematic documentation of acceptable security degradations, emergency
access mechanisms, and security debt tracking during disruptive events.

**Generated Workbook Structure:**
1. Instructions - Assessment guidance and methodology
2. Degradation_Scenarios - Acceptable degradation documentation
3. BreakGlass_Accounts - Emergency account inventory
4. BreakGlass_Activation - Activation log and procedures
5. Elevated_Monitoring - Enhanced monitoring requirements
6. Personnel_Availability - Security team succession
7. Security_Debt_Register - Deferred security tracking
8. Evidence_Register - Audit evidence tracking
9. Approval_SignOff - Stakeholder review and approval

================================================================================
"""

# =============================================================================
# IMPORTS - Standard Library
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# IMPORTS - Third Party
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.29.2"
WORKBOOK_NAME = "Degraded Mode Security Requirements"
CONTROL_ID = "A.5.29"
CONTROL_NAME = "Information Security During Disruption"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "warning": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Degradation_Scenarios",
        "BreakGlass_Accounts",
        "BreakGlass_Activation",
        "Elevated_Monitoring",
        "Personnel_Availability",
        "Security_Debt_Register",
        "Evidence_Register",
        "Approval_SignOff",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Degraded Mode Security Requirements"),
        ("Control Reference", CONTROL_ID),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Annual"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Document acceptable Degradation_Scenarios with compensating controls",
        "2. Inventory all BreakGlass_Accounts with activation procedures",
        "3. Use BreakGlass_Activation log for any emergency access activations",
        "4. Define Elevated_Monitoring requirements per posture level",
        "5. Document Personnel_Availability with succession planning",
        "6. Track Security_Debt_Register items requiring remediation",
        "7. Link evidence in Evidence_Register",
        "8. Obtain approvals in Approval_SignOff sheet",
    ]

    row += 1
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    ws.freeze_panes = "A4"


# =============================================================================
# DEGRADATION SCENARIOS SHEET
# =============================================================================
def create_degradation_scenarios_sheet(ws, styles):
    """Create the Degradation_Scenarios sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "DEGRADATION SCENARIOS - Acceptable Security Relaxations"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Scenario_ID", 15),
        ("Scenario_Name", 30),
        ("Trigger_Condition", 40),
        ("Control_Affected", 25),
        ("Degradation_Type", 22),
        ("Compensating_Control", 45),
        ("Max_Duration", 20),
        ("Renewal_Process", 35),
        ("Authorisation_Required", 25),
        ("Posture_Level", 15),
        ("Never_Acceptable", 40),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Sample scenarios
    scenarios = [
        ("DS-001", "MFA Infrastructure Unavailable",
         "MFA platform failure or inaccessible",
         "Multi-Factor Authentication",
         "Temporary Bypass",
         "Enhanced logging, IP restrictions, session limits, 4-hour max sessions",
         "72 hours", "CISO approval required for each 24h extension",
         "CISO", "Degraded",
         "Allowing permanent single-factor access"),
        ("DS-002", "SIEM Platform Unavailable",
         "SIEM system failure or maintenance",
         "Security Monitoring",
         "Reduced Capability",
         "Manual log review every 4 hours for critical systems",
         "48 hours", "Security Manager approval for extension",
         "Security Manager", "Degraded",
         "No logging of security events"),
        ("DS-003", "Network Segmentation Bypass",
         "Emergency connectivity requirement",
         "Network Segmentation",
         "Temporary Bypass",
         "Enhanced monitoring, specific ports only, time-limited",
         "24 hours", "CIO + CISO approval for extension",
         "CIO + CISO", "Emergency",
         "Permanent network flattening"),
        ("DS-004", "Patch Management Delayed",
         "System stability during disruption",
         "Patch Management",
         "Postponed",
         "Critical patches only, manual review process",
         "30 days non-critical", "CISO approval for extension",
         "CISO", "Degraded",
         "Ignoring critical security patches"),
        ("DS-005", "Access Review Postponed",
         "Resource constraints during disruption",
         "Access Reviews",
         "Postponed",
         "Stricter approval for new access, no new privileged access",
         "30 days", "Security Manager approval for extension",
         "Security Manager", "Elevated",
         "Granting new privileged access without review"),
    ]

    dv_type = DataValidation(
        type="list",
        formula1='"Temporary Bypass,Reduced Capability,Postponed,Alternative Method"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_posture = DataValidation(
        type="list",
        formula1='"Elevated,Degraded,Emergency"',
        allow_blank=False
    )
    ws.add_data_validation(dv_posture)

    for row_idx, scenario in enumerate(scenarios, start=4):
        for col_idx, value in enumerate(scenario, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            if col_idx in [3, 6, 8, 11]:
                cell.alignment = Alignment(wrap_text=True, vertical="center")

        dv_type.add(ws.cell(row=row_idx, column=5))
        dv_posture.add(ws.cell(row=row_idx, column=10))

    for r in range(len(scenarios) + 4, len(scenarios) + 24):
        ws.cell(row=r, column=1, value=f"DS-{r-3:03d}").font = Font(color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_type.add(ws.cell(row=r, column=5))
        dv_posture.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "C4"


# =============================================================================
# BREAK-GLASS ACCOUNTS SHEET
# =============================================================================
def create_breakglass_accounts_sheet(ws, styles):
    """Create the BreakGlass_Accounts sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = "BREAK-GLASS ACCOUNTS - Emergency Access Inventory"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Account_ID", 15),
        ("Account_Name", 25),
        ("Account_Type", 20),
        ("Target_Systems", 35),
        ("Scope_Permissions", 35),
        ("Credential_Location", 30),
        ("Activation_Authority", 25),
        ("Two_Person_Required", 15),
        ("Default_Duration", 15),
        ("Logging_Enabled", 15),
        ("Last_Rotation_Date", 16),
        ("Last_Test_Date", 16),
        ("Status", 12),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Domain Admin,System Admin,Database Admin,Network Admin,Application Admin,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_bool = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_bool)

    dv_logging = DataValidation(
        type="list",
        formula1='"Yes,No,Verified"',
        allow_blank=False
    )
    ws.add_data_validation(dv_logging)

    dv_status = DataValidation(
        type="list",
        formula1='"Disabled,Enabled,Unknown"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 24):
        ws.cell(row=r, column=1, value=f"BG-{r-3:03d}").font = Font(color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        dv_type.add(ws.cell(row=r, column=3))
        dv_bool.add(ws.cell(row=r, column=8))
        dv_logging.add(ws.cell(row=r, column=10))
        dv_status.add(ws.cell(row=r, column=13))

    ws.freeze_panes = "C4"


# =============================================================================
# BREAK-GLASS ACTIVATION SHEET
# =============================================================================
def create_breakglass_activation_sheet(ws, styles):
    """Create the BreakGlass_Activation sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = "BREAK-GLASS ACTIVATION LOG"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Activation_ID", 15),
        ("Account_ID", 15),
        ("Emergency_Type", 25),
        ("Activation_DateTime", 20),
        ("Authorised_By", 25),
        ("Activated_By", 25),
        ("Second_Person", 25),
        ("CISO_Notified", 12),
        ("Expiry_DateTime", 20),
        ("Renewed", 10),
        ("Deactivation_DateTime", 20),
        ("Post_Review_Completed", 15),
        ("Issues_Found", 40),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_bool = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_bool)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"ACT-{r-3:03d}").font = Font(color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        dv_bool.add(ws.cell(row=r, column=8))
        dv_bool.add(ws.cell(row=r, column=10))
        dv_bool.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "C4"


# =============================================================================
# ELEVATED MONITORING SHEET
# =============================================================================
def create_elevated_monitoring_sheet(ws, styles):
    """Create the Elevated_Monitoring sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "ELEVATED MONITORING - Enhanced Surveillance Requirements"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Posture_Level", 15),
        ("Monitoring_Area", 25),
        ("Normal_Frequency", 20),
        ("Enhanced_Frequency", 20),
        ("Alert_Threshold_Change", 35),
        ("Manual_Review_Required", 15),
        ("Review_Frequency", 20),
        ("Responsible_Party", 25),
        ("Implementation_Notes", 40),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate with monitoring requirements
    monitoring = [
        ("Elevated", "Authentication Logs", "Hourly", "15 minutes",
         "Lower failed attempt threshold from 5 to 3", "No", "", "SOC", ""),
        ("Elevated", "Privileged Access", "Daily", "Hourly",
         "Alert on any privileged access", "Yes", "Every 4 hours", "Security Team", ""),
        ("Degraded", "Authentication Logs", "15 minutes", "Real-time",
         "Alert on any anomaly", "Yes", "Every 2 hours", "SOC", "24/7 coverage required"),
        ("Degraded", "Network Traffic", "Hourly", "15 minutes",
         "Alert on unusual destinations", "Yes", "Every 4 hours", "Network Team", "Focus on egress"),
        ("Emergency", "All Critical Systems", "Real-time", "Real-time",
         "Maximum sensitivity", "Yes", "Continuous", "SOC + Security Team", "Threat hunting focus"),
        ("Emergency", "Break-Glass Usage", "N/A", "Real-time",
         "Immediate alert on any activation", "Yes", "Continuous", "CISO", "Direct notification"),
    ]

    dv_posture = DataValidation(
        type="list",
        formula1='"Elevated,Degraded,Emergency"',
        allow_blank=False
    )
    ws.add_data_validation(dv_posture)

    dv_bool = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_bool)

    for row_idx, mon in enumerate(monitoring, start=4):
        for col_idx, value in enumerate(mon, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            if col_idx in [5, 9]:
                cell.alignment = Alignment(wrap_text=True, vertical="center")

        dv_posture.add(ws.cell(row=row_idx, column=1))
        dv_bool.add(ws.cell(row=row_idx, column=6))

    for r in range(len(monitoring) + 4, len(monitoring) + 24):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_posture.add(ws.cell(row=r, column=1))
        dv_bool.add(ws.cell(row=r, column=6))

    ws.freeze_panes = "C4"


# =============================================================================
# PERSONNEL AVAILABILITY SHEET
# =============================================================================
def create_personnel_availability_sheet(ws, styles):
    """Create the Personnel_Availability sheet."""
    ws.merge_cells("A1:N1")
    ws["A1"] = "PERSONNEL AVAILABILITY - Security Team Succession"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Role_ID", 12),
        ("Role_Name", 30),
        ("Primary_Name", 25),
        ("Primary_Phone", 20),
        ("Primary_Email", 30),
        ("Backup1_Name", 25),
        ("Backup1_Phone", 20),
        ("Backup1_Email", 30),
        ("Backup2_Name", 25),
        ("Backup2_Phone", 20),
        ("Backup2_Email", 30),
        ("Cross_Training_Status", 18),
        ("Last_Contact_Test", 16),
        ("Last_Drill_Date", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate with key roles
    roles = [
        ("R-001", "CISO"),
        ("R-002", "Security Manager"),
        ("R-003", "Security Analyst (Lead)"),
        ("R-004", "SOC Lead"),
        ("R-005", "Incident Response Lead"),
        ("R-006", "IAM Administrator"),
        ("R-007", "Network Security Lead"),
        ("R-008", "Security Architect"),
    ]

    dv_training = DataValidation(
        type="list",
        formula1='"Complete,Partial,None"',
        allow_blank=False
    )
    ws.add_data_validation(dv_training)

    for row_idx, (role_id, role_name) in enumerate(roles, start=4):
        ws.cell(row=row_idx, column=1, value=role_id)
        ws.cell(row=row_idx, column=2, value=role_name)
        for c in range(3, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_training.add(ws.cell(row=row_idx, column=12))

    for r in range(len(roles) + 4, len(roles) + 14):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_training.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "C4"


# =============================================================================
# SECURITY DEBT REGISTER SHEET
# =============================================================================
def create_security_debt_register_sheet(ws, styles):
    """Create the Security_Debt_Register sheet."""
    ws.merge_cells("A1:O1")
    ws["A1"] = "SECURITY DEBT REGISTER - Deferred Security Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Debt_ID", 12),
        ("Debt_Type", 20),
        ("Description", 45),
        ("Disruption_Reference", 20),
        ("Created_Date", 16),
        ("Owner", 25),
        ("Remediation_Plan", 45),
        ("Target_Date", 16),
        ("Status", 15),
        ("Age_Days", 12),
        ("Escalation_Required", 15),
        ("Escalated_To", 25),
        ("Escalation_Date", 16),
        ("Closure_Date", 16),
        ("Closure_Evidence", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Deferred Patch,Skipped Review,Delayed Scan,Config Exception,Access Exception,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Closed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"SD-{r-3:03d}").font = Font(color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        dv_type.add(ws.cell(row=r, column=2))
        dv_status.add(ws.cell(row=r, column=9))

        # Age formula
        ws.cell(row=r, column=10).value = f'=IF(E{r}<>"",TODAY()-E{r},"")'
        ws.cell(row=r, column=10).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Escalation formula
        ws.cell(row=r, column=11).value = f'=IF(J{r}>30,"Yes","No")'
        ws.cell(row=r, column=11).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.freeze_panes = "D4"


# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================
def create_evidence_register_sheet(ws, styles):
    """Create the Evidence_Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 45),
        ("Related_Section", 25),
        ("Collection_Date", 16),
        ("Location", 40),
        ("Collected_By", 25),
        ("Valid_Until", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Document,Approval,Test Result,Configuration,Screenshot,Attestation"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "C4"


# =============================================================================
# APPROVAL SIGN-OFF SHEET
# =============================================================================
def create_approval_signoff_sheet(ws, styles):
    """Create the Approval_SignOff sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Total Break-Glass Accounts", "=COUNTA(BreakGlass_Accounts!A4:A23)-COUNTBLANK(BreakGlass_Accounts!B4:B23)"),
        ("Accounts Tested This Year", '=COUNTIF(BreakGlass_Accounts!L4:L23,">="&DATE(YEAR(TODAY()),1,1))'),
        ("Open Security Debt Items", '=COUNTIF(Security_Debt_Register!I4:I53,"Open")'),
        ("Overdue Security Debt (>30 days)", '=COUNTIF(Security_Debt_Register!K4:K53,"Yes")'),
        ("Assessment Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "" or not str(value).startswith("="):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    for field in ["Name", "Role/Title", "Department", "Email", "Date"]:
        row += 1
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (Security Manager)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    for field in ["Name", "Date", "Signature"]:
        row += 1
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_decision)

    for field in ["Name", "Date", "Approval Decision", "Signature"]:
        row += 1
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        if field == "Approval Decision":
            dv_decision.add(ws[f"B{row}"])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A3"


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/9] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/9] Creating Degradation_Scenarios sheet...")
        create_degradation_scenarios_sheet(wb["Degradation_Scenarios"], styles)

        logger.info("[3/9] Creating BreakGlass_Accounts sheet...")
        create_breakglass_accounts_sheet(wb["BreakGlass_Accounts"], styles)

        logger.info("[4/9] Creating BreakGlass_Activation sheet...")
        create_breakglass_activation_sheet(wb["BreakGlass_Activation"], styles)

        logger.info("[5/9] Creating Elevated_Monitoring sheet...")
        create_elevated_monitoring_sheet(wb["Elevated_Monitoring"], styles)

        logger.info("[6/9] Creating Personnel_Availability sheet...")
        create_personnel_availability_sheet(wb["Personnel_Availability"], styles)

        logger.info("[7/9] Creating Security_Debt_Register sheet...")
        create_security_debt_register_sheet(wb["Security_Debt_Register"], styles)

        logger.info("[8/9] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[9/9] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        logger.error("Install with: pip install openpyxl")
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.29.2 specification
# =============================================================================
