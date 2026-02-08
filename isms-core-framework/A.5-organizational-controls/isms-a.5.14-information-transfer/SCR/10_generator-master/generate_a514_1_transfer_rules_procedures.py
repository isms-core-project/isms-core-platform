#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
A.5.14.1 Transfer Rules and Procedures Generator
================================================================================

Generates Excel workbook for documenting information transfer rules, procedures,
and acceptable methods based on classification levels per ISO 27001:2022 A.5.14.

Sheets:
    1. Instructions - Completion guidance
    2. Transfer_Policy - Overarching transfer policy elements
    3. Transfer_Methods - Acceptable methods by classification level
    4. Electronic_Transfer - Email, cloud, messaging rules
    5. Physical_Transfer - Media, courier, document rules
    6. Verbal_Transfer - Meeting, call, discussion protocols
    7. Evidence_Register - Supporting documentation
    8. Approval_SignOff - Authorization workflow

Usage:
    python3 generate_a514_1_transfer_rules_procedures.py

Output:
    ISMS-IMP-A.5.14.1_Transfer_Rules_and_Procedures_YYYYMMDD.xlsx

================================================================================
"""

import logging
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.14.1"
WORKBOOK_NAME = "Transfer Rules and Procedures"
CONTROL_ID = "A.5.14"
CONTROL_NAME = "Information Transfer"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
# Colors
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
INSTRUCTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Classification colors
PUBLIC_FILL = PatternFill(start_color="74C0FC", end_color="74C0FC", fill_type="solid")
INTERNAL_FILL = PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")
CONFIDENTIAL_FILL = PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid")
RESTRICTED_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

# Fonts
HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)
INSTRUCTION_FONT = Font(name="Calibri", size=10, italic=True)

# Borders
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Alignment
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP_LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================
def set_column_widths(ws, widths: dict):
    """Set column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_header_row(ws, row: int, headers: list, fill=SUBHEADER_FILL, font=SUBHEADER_FONT):
    """Create a styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def add_data_validation(ws, cell_range: str, formula: str, title: str = "", prompt: str = ""):
    """Add dropdown data validation."""
    dv = DataValidation(
        type="list",
        formula1=formula,
        showDropDown=False,
        allowBlank=True
    )
    if title:
        dv.promptTitle = title
    if prompt:
        dv.prompt = prompt
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


# =============================================================================
# SHEET CREATORS
# =============================================================================
def create_instructions_sheet(wb: Workbook):
    """Create the Instructions sheet."""
    ws = wb.active
    ws.title = "Instructions"

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Metadata
    metadata = [
        ("Document ID:", DOCUMENT_ID),
        ("Control Reference:", CONTROL_REF),
        ("Generated Date:", GENERATED_DATE),
        ("Version:", "1.0"),
        ("Classification:", "INTERNAL"),
    ]

    row = 3
    for label, value in metadata:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        row += 1

    # Purpose
    row += 1
    ws.cell(row=row, column=1, value="PURPOSE").font = BOLD_FONT
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    purpose_cell = ws.cell(row=row, column=1)
    purpose_cell.value = (
        "This workbook documents information transfer rules and procedures per ISO 27001:2022 Control A.5.14. "
        "It defines acceptable transfer methods for each classification level, covering electronic, physical, "
        "and verbal information exchanges. Properly completed, this workbook demonstrates compliance with "
        "requirements for protecting information during transfer."
    )
    purpose_cell.font = NORMAL_FONT
    purpose_cell.alignment = TOP_LEFT_ALIGN
    ws.row_dimensions[row].height = 60

    # Sheet descriptions
    row += 2
    ws.cell(row=row, column=1, value="SHEET DESCRIPTIONS").font = BOLD_FONT
    row += 1

    sheets = [
        ("Transfer_Policy", "Overarching policy elements for information transfer"),
        ("Transfer_Methods", "Acceptable transfer methods mapped to classification levels"),
        ("Electronic_Transfer", "Rules for email, cloud services, messaging, and file sharing"),
        ("Physical_Transfer", "Procedures for media, courier, and document transfers"),
        ("Verbal_Transfer", "Protocols for meetings, calls, and verbal discussions"),
        ("Evidence_Register", "Supporting documentation and evidence tracking"),
        ("Approval_SignOff", "Authorization and approval workflow"),
    ]

    for sheet_name, description in sheets:
        ws.cell(row=row, column=1, value=sheet_name).font = BOLD_FONT
        ws.cell(row=row, column=2, value=description).font = NORMAL_FONT
        row += 1

    # Color legend
    row += 1
    ws.cell(row=row, column=1, value="CLASSIFICATION COLOR CODING").font = BOLD_FONT
    row += 1

    colors = [
        ("PUBLIC", PUBLIC_FILL, "Information approved for public release"),
        ("INTERNAL", INTERNAL_FILL, "Internal business information"),
        ("CONFIDENTIAL", CONFIDENTIAL_FILL, "Sensitive business information"),
        ("RESTRICTED", RESTRICTED_FILL, "Highly sensitive, limited access"),
    ]

    for level, fill, desc in colors:
        cell = ws.cell(row=row, column=1, value=level)
        cell.fill = fill
        cell.font = BOLD_FONT
        ws.cell(row=row, column=2, value=desc).font = NORMAL_FONT
        row += 1

    set_column_widths(ws, {"A": 25, "B": 60, "C": 20, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20})
    ws.freeze_panes = "A3"

    logger.info("Created Instructions sheet")


def create_transfer_policy_sheet(wb: Workbook):
    """Create the Transfer Policy sheet."""
    ws = wb.create_sheet("Transfer_Policy")

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Information Transfer Policy Elements"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Policy Element", "Description", "Requirement", "Owner", "Review Frequency", "Status"]
    create_header_row(ws, 3, headers)

    # Policy elements
    policies = [
        ("Scope Definition", "Define what information transfer means in organizational context", "MANDATORY", "", "Annual", ""),
        ("Classification Alignment", "Transfer methods must align with information classification", "MANDATORY", "", "Annual", ""),
        ("Encryption Standards", "Minimum encryption requirements for transfer by classification", "MANDATORY", "", "Annual", ""),
        ("Authentication Requirements", "Recipient verification before transfer", "MANDATORY", "", "Annual", ""),
        ("Authorization Process", "Approval workflow for transfers above certain thresholds", "MANDATORY", "", "Annual", ""),
        ("Chain of Custody", "Tracking requirements for information in transit", "RECOMMENDED", "", "Annual", ""),
        ("Third-Party Requirements", "Standards for transfers to/from external parties", "MANDATORY", "", "Annual", ""),
        ("Incident Reporting", "Process for reporting transfer security incidents", "MANDATORY", "", "Annual", ""),
        ("Retention of Transfer Records", "How long transfer logs must be retained", "MANDATORY", "", "Annual", ""),
        ("User Awareness", "Training requirements for secure transfer practices", "MANDATORY", "", "Annual", ""),
    ]

    row = 4
    for policy in policies:
        for col, value in enumerate(policy, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col == 4 or col == 6:  # Input fields
                cell.fill = INPUT_FILL
        row += 1

    # Data validation for Status
    add_data_validation(ws, f"F4:F{row-1}", '"Draft,Under Review,Approved,Needs Update"')

    set_column_widths(ws, {"A": 30, "B": 50, "C": 15, "D": 25, "E": 18, "F": 15})
    ws.freeze_panes = "A4"

    logger.info("Created Transfer_Policy sheet")


def create_transfer_methods_sheet(wb: Workbook):
    """Create the Transfer Methods matrix sheet."""
    ws = wb.create_sheet("Transfer_Methods")

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Acceptable Transfer Methods by Classification Level"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Transfer Method", "PUBLIC", "INTERNAL", "CONFIDENTIAL", "RESTRICTED", "Conditions", "Approval Required", "Notes"]
    create_header_row(ws, 3, headers)

    # Apply classification colors to headers
    ws["B3"].fill = PUBLIC_FILL
    ws["C3"].fill = INTERNAL_FILL
    ws["D3"].fill = CONFIDENTIAL_FILL
    ws["E3"].fill = RESTRICTED_FILL

    # Transfer methods
    methods = [
        ("Corporate Email (encrypted)", "Allowed", "Allowed", "Allowed", "Not Allowed", "TLS required", "No", ""),
        ("Corporate Email (unencrypted)", "Allowed", "Conditional", "Not Allowed", "Not Allowed", "Internal only", "No", ""),
        ("Personal Email", "Conditional", "Not Allowed", "Not Allowed", "Not Allowed", "Emergency only", "Yes", ""),
        ("Approved Cloud Storage", "Allowed", "Allowed", "Allowed", "Conditional", "Encrypted, access controlled", "Varies", ""),
        ("Public Cloud Storage", "Allowed", "Not Allowed", "Not Allowed", "Not Allowed", "Non-sensitive only", "No", ""),
        ("SFTP/SCP", "Allowed", "Allowed", "Allowed", "Allowed", "Key-based auth", "No", ""),
        ("FTP (unencrypted)", "Conditional", "Not Allowed", "Not Allowed", "Not Allowed", "Isolated networks", "Yes", ""),
        ("USB/Removable Media", "Conditional", "Conditional", "Conditional", "Not Allowed", "Encrypted media only", "Yes", ""),
        ("Physical Courier", "Allowed", "Allowed", "Allowed", "Allowed", "Tamper-evident packaging", "Varies", ""),
        ("Internal Mail", "Allowed", "Allowed", "Conditional", "Not Allowed", "Sealed envelopes", "No", ""),
        ("Video Conference", "Allowed", "Allowed", "Conditional", "Conditional", "Approved platforms", "No", ""),
        ("Telephone", "Allowed", "Allowed", "Conditional", "Not Allowed", "Verify recipient", "No", ""),
        ("Instant Messaging (approved)", "Allowed", "Allowed", "Conditional", "Not Allowed", "Corporate tools only", "No", ""),
        ("Instant Messaging (public)", "Conditional", "Not Allowed", "Not Allowed", "Not Allowed", "Non-business", "Yes", ""),
        ("API/Web Services", "Allowed", "Allowed", "Allowed", "Conditional", "mTLS, OAuth2", "Technical Review", ""),
    ]

    row = 4
    for method in methods:
        for col, value in enumerate(method, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if col in [2, 3, 4, 5, 7] else LEFT_ALIGN
            # Color code based on value
            if value == "Allowed":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif value == "Conditional":
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif value == "Not Allowed":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 1

    # Data validation
    add_data_validation(ws, f"B4:E{row+10}", '"Allowed,Conditional,Not Allowed"')
    add_data_validation(ws, f"G4:G{row+10}", '"Yes,No,Varies,Technical Review"')

    set_column_widths(ws, {"A": 30, "B": 12, "C": 12, "D": 15, "E": 12, "F": 30, "G": 18, "H": 30})
    ws.freeze_panes = "A4"

    logger.info("Created Transfer_Methods sheet")


def create_electronic_transfer_sheet(wb: Workbook):
    """Create the Electronic Transfer rules sheet."""
    ws = wb.create_sheet("Electronic_Transfer")

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Electronic Transfer Rules and Controls"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Channel", "Security Controls", "Encryption Requirement", "Authentication", "Logging", "Max Classification", "DLP Integration", "Notes"]
    create_header_row(ws, 3, headers)

    # Electronic channels
    channels = [
        ("Corporate Email - Internal", "Exchange Online, Microsoft 365", "TLS 1.2+ in transit", "Microsoft Entra ID (formerly Azure AD) SSO", "Exchange Audit Logs", "CONFIDENTIAL", "Yes", ""),
        ("Corporate Email - External", "Exchange Online + S/MIME", "S/MIME or TLS 1.2+", "Microsoft Entra ID (formerly Azure AD) + MFA", "Exchange Audit Logs", "CONFIDENTIAL", "Yes", ""),
        ("SharePoint/OneDrive", "Microsoft 365, Sensitivity Labels", "AES-256 at rest, TLS in transit", "Microsoft Entra ID (formerly Azure AD) + MFA", "Unified Audit Log", "CONFIDENTIAL", "Yes", ""),
        ("Teams File Sharing", "Microsoft Teams", "AES-256, TLS 1.2+", "Microsoft Entra ID (formerly Azure AD) + MFA", "Teams Audit", "INTERNAL", "Yes", ""),
        ("SFTP Server", "OpenSSH, key-based auth", "AES-256-CTR", "SSH Keys + IP allowlist", "SFTP Logs", "RESTRICTED", "Manual", ""),
        ("Approved API Gateway", "Kong/Apigee with mTLS", "TLS 1.3, mTLS", "OAuth 2.0 + API Keys", "API Gateway Logs", "RESTRICTED", "Custom", ""),
        ("Managed File Transfer (MFT)", "GoAnywhere/Axway", "AES-256, TLS 1.2+", "SSO + MFA", "MFT Audit Trail", "RESTRICTED", "Yes", ""),
        ("VPN Site-to-Site", "IPSec/IKEv2", "AES-256-GCM", "Certificate + PSK", "VPN Logs", "RESTRICTED", "N/A", ""),
    ]

    row = 4
    for channel in channels:
        for col, value in enumerate(channel, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col == 8:  # Notes input
                cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"F4:F{row+5}", '"PUBLIC,INTERNAL,CONFIDENTIAL,RESTRICTED"')
    add_data_validation(ws, f"G4:G{row+5}", '"Yes,No,Manual,Custom,N/A"')

    set_column_widths(ws, {"A": 28, "B": 30, "C": 25, "D": 22, "E": 20, "F": 18, "G": 15, "H": 30})
    ws.freeze_panes = "A4"

    logger.info("Created Electronic_Transfer sheet")


def create_physical_transfer_sheet(wb: Workbook):
    """Create the Physical Transfer procedures sheet."""
    ws = wb.create_sheet("Physical_Transfer")

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Physical Transfer Procedures and Controls"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Transfer Type", "Packaging Requirements", "Courier/Transport", "Chain of Custody", "Recipient Verification", "Max Classification", "Environmental Protection", "Notes"]
    create_header_row(ws, 3, headers)

    # Physical transfer types
    transfers = [
        ("Encrypted USB Drive", "Tamper-evident bag", "Hand delivery or approved courier", "Transfer log required", "ID verification + signature", "CONFIDENTIAL", "Static protection", ""),
        ("External Hard Drive", "Padded, tamper-evident case", "Approved courier with tracking", "Full chain of custody log", "ID verification + signature", "CONFIDENTIAL", "Static + shock protection", ""),
        ("Printed Documents", "Sealed envelope/package", "Internal mail or courier", "Receipt confirmation", "Addressee verification", "INTERNAL", "Waterproof packaging", ""),
        ("Classified Documents", "Double-sealed, opaque packaging", "Approved secure courier", "Continuous chain of custody", "Photo ID + signature + witness", "RESTRICTED", "Full environmental protection", ""),
        ("Backup Tapes", "Protective case, sealed", "Secure transport service", "Barcode tracking", "Authorized recipient list", "RESTRICTED", "Temperature controlled", ""),
        ("Mobile Devices", "Original or protective packaging", "Hand delivery preferred", "Asset tracking log", "Employee ID verification", "CONFIDENTIAL", "Impact protection", ""),
        ("Hardware for Disposal", "Secure container", "Certified disposal vendor", "Destruction certificate", "Vendor certification", "RESTRICTED", "Secure transport", ""),
    ]

    row = 4
    for transfer in transfers:
        for col, value in enumerate(transfer, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col == 8:  # Notes input
                cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"F4:F{row+5}", '"PUBLIC,INTERNAL,CONFIDENTIAL,RESTRICTED"')

    set_column_widths(ws, {"A": 25, "B": 28, "C": 30, "D": 25, "E": 28, "F": 18, "G": 25, "H": 30})
    ws.freeze_panes = "A4"

    logger.info("Created Physical_Transfer sheet")


def create_verbal_transfer_sheet(wb: Workbook):
    """Create the Verbal Transfer protocols sheet."""
    ws = wb.create_sheet("Verbal_Transfer")

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "Verbal Transfer Protocols"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Communication Type", "Security Precautions", "Participant Verification", "Recording Policy", "Max Classification", "Follow-up Documentation", "Notes"]
    create_header_row(ws, 3, headers)

    # Verbal transfer types
    verbal = [
        ("In-Person Meeting (internal)", "Secure meeting room, no unauthorized persons", "Badge/employee verification", "No recording without consent", "CONFIDENTIAL", "Meeting minutes if needed", ""),
        ("In-Person Meeting (external)", "Visitor registration, escort required", "ID verification, NDA check", "Recording requires approval", "INTERNAL", "Summary + attendee list", ""),
        ("Video Conference (internal)", "Approved platform (Teams/Zoom)", "Microsoft Entra ID (formerly Azure AD) authentication", "Recording notification required", "CONFIDENTIAL", "Auto-transcription available", ""),
        ("Video Conference (external)", "Lobby/waiting room enabled", "Host verification of attendees", "Explicit consent required", "INTERNAL", "Host controls recording", ""),
        ("Phone Call (internal)", "Corporate phone system", "Caller ID verification", "Generally not recorded", "INTERNAL", "Notes if sensitive", ""),
        ("Phone Call (external)", "Verify callback number if sensitive", "Ask security questions", "Inform if recorded", "PUBLIC", "Document key points", ""),
        ("Voicemail", "PIN-protected access", "N/A", "Limited retention period", "PUBLIC", "Transcription if available", ""),
        ("Presentation/Training", "Room access control", "Attendee list required", "Prior approval needed", "INTERNAL", "Presentation materials archived", ""),
    ]

    row = 4
    for item in verbal:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col == 7:  # Notes input
                cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"E4:E{row+5}", '"PUBLIC,INTERNAL,CONFIDENTIAL,RESTRICTED"')

    set_column_widths(ws, {"A": 30, "B": 35, "C": 28, "D": 28, "E": 18, "F": 28, "G": 30})
    ws.freeze_panes = "A4"

    logger.info("Created Verbal_Transfer sheet")


def create_evidence_register_sheet(wb: Workbook):
    """Create the Evidence Register sheet."""
    ws = wb.create_sheet("Evidence_Register")

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Evidence Register - Information Transfer Controls"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Evidence ID", "Evidence Type", "Description", "Related Control", "Location/Link", "Date Collected", "Collected By", "Status"]
    create_header_row(ws, 3, headers)

    # Sample evidence entries
    evidence = [
        ("EV-514-001", "Policy Document", "Information Transfer Policy", "Transfer_Policy", "", "", "", ""),
        ("EV-514-002", "Configuration Export", "Email encryption settings", "Electronic_Transfer", "", "", "", ""),
        ("EV-514-003", "Procedure Document", "Secure courier procedures", "Physical_Transfer", "", "", "", ""),
        ("EV-514-004", "Training Record", "Transfer security awareness completion", "Verbal_Transfer", "", "", "", ""),
        ("EV-514-005", "Audit Log Sample", "30-day transfer audit log extract", "Electronic_Transfer", "", "", "", ""),
    ]

    row = 4
    for item in evidence:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [5, 6, 7, 8]:  # Input fields
                cell.fill = INPUT_FILL
        row += 1

    # Add empty rows for additional entries
    for i in range(10):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"H4:H{row-1}", '"Pending,Collected,Verified,Expired,N/A"')
    add_data_validation(ws, f"B4:B{row-1}", '"Policy Document,Procedure Document,Configuration Export,Audit Log Sample,Training Record,Screenshot,Agreement,Other"')

    set_column_widths(ws, {"A": 15, "B": 20, "C": 35, "D": 20, "E": 30, "F": 15, "G": 18, "H": 12})
    ws.freeze_panes = "A4"

    logger.info("Created Evidence_Register sheet")


def create_approval_signoff_sheet(wb: Workbook):
    """Create the Approval and Sign-Off sheet."""
    ws = wb.create_sheet("Approval_SignOff")

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Transfer Rules Approval and Sign-Off"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Document info section
    ws.cell(row=3, column=1, value="Document Information").font = BOLD_FONT
    ws.merge_cells("A3:F3")

    info_rows = [
        ("Document ID:", DOCUMENT_ID, "Version:", "1.0"),
        ("Document Title:", WORKBOOK_NAME, "Date:", GENERATED_DATE),
        ("Control Reference:", CONTROL_ID, "Classification:", "INTERNAL"),
    ]

    row = 4
    for item in info_rows:
        ws.cell(row=row, column=1, value=item[0]).font = BOLD_FONT
        ws.cell(row=row, column=2, value=item[1]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=item[2]).font = BOLD_FONT
        ws.cell(row=row, column=4, value=item[3]).font = NORMAL_FONT
        row += 1

    # Approval section
    row += 1
    ws.cell(row=row, column=1, value="Approval Signatures").font = BOLD_FONT
    ws.merge_cells(f"A{row}:F{row}")

    row += 1
    approval_headers = ["Role", "Name", "Signature", "Date", "Status", "Comments"]
    create_header_row(ws, row, approval_headers)

    row += 1
    approvers = [
        ("Document Owner", "", "", "", "", ""),
        ("IT Security Manager", "", "", "", "", ""),
        ("Information Security Officer", "", "", "", "", ""),
        ("Department Head", "", "", "", "", ""),
        ("CISO (if RESTRICTED)", "", "", "", "", ""),
    ]

    for approver in approvers:
        for col, value in enumerate(approver, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col > 1:  # Input fields
                cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"E{row-5}:E{row-1}", '"Pending,Approved,Rejected,Deferred"')

    # Version history section
    row += 2
    ws.cell(row=row, column=1, value="Version History").font = BOLD_FONT
    ws.merge_cells(f"A{row}:F{row}")

    row += 1
    version_headers = ["Version", "Date", "Author", "Changes", "Approved By", ""]
    create_header_row(ws, row, version_headers)

    row += 1
    ws.cell(row=row, column=1, value="1.0").font = NORMAL_FONT
    ws.cell(row=row, column=2, value=GENERATED_DATE).font = NORMAL_FONT
    ws.cell(row=row, column=3, value="").font = NORMAL_FONT
    ws.cell(row=row, column=3).fill = INPUT_FILL
    ws.cell(row=row, column=4, value="Initial version").font = NORMAL_FONT
    ws.cell(row=row, column=5, value="").font = NORMAL_FONT
    ws.cell(row=row, column=5).fill = INPUT_FILL

    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER

    set_column_widths(ws, {"A": 25, "B": 25, "C": 20, "D": 15, "E": 15, "F": 30})
    ws.freeze_panes = "A4"

    logger.info("Created Approval_SignOff sheet")


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main():
    """Main execution function."""
    logger.info("=" * 70)
    logger.info(f"{DOCUMENT_ID} {WORKBOOK_NAME} Generator")
    logger.info("=" * 70)

    # Create workbook
    wb = Workbook()

    # Create all sheets
    create_instructions_sheet(wb)
    create_transfer_policy_sheet(wb)
    create_transfer_methods_sheet(wb)
    create_electronic_transfer_sheet(wb)
    create_physical_transfer_sheet(wb)
    create_verbal_transfer_sheet(wb)
    create_evidence_register_sheet(wb)
    create_approval_signoff_sheet(wb)

    # Save workbook
    output_path = Path(__file__).parent / OUTPUT_FILENAME
    wb.save(output_path)

    logger.info("=" * 70)
    logger.info(f"SUCCESS: Workbook saved as {OUTPUT_FILENAME}")
    logger.info("=" * 70)

    return 0


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.14 Information Transfer control
# =============================================================================
