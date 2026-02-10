#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.32-33.S3 - Retention and Disposal Schedule Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.33: Protection of Records
"Records should be protected from loss, destruction, falsification, unauthorised
access and unauthorised release in accordance with legislative, regulatory,
contractual and business requirements."

Assessment Domain 3 of 4: Retention and Disposal Schedule

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel workbook for managing record
retention periods and secure disposal processes.

**Purpose:**
Establish retention requirements for all record categories and manage
the secure disposal of records when retention periods expire.

**Generated Workbook Structure:**
1. Instructions - Guidance on retention and disposal management
2. Retention_Schedule - Retention periods for all record categories
3. Regulatory_Mapping - Regulations driving retention requirements
4. Disposal_Queue - Records due for disposal
5. Disposal_Method_Matrix - Appropriate methods per classification
6. Destruction_Verification - Destruction evidence tracking
7. Exception_Register - Retention extensions and early disposals
8. Compliance_Dashboard - Schedule adherence metrics
9. Gap_Analysis - Issues and remediation
10. Evidence_Register - Audit evidence tracking
11. Approval_SignOff - Assessment approval

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.32-33.S3"
WORKBOOK_NAME = "Retention Disposal Schedule"
CONTROL_ID = "A.5.33"
CONTROL_NAME = "Protection of Records"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "overdue": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "approaching": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "on_track": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    }


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    for attr, value in style_dict.items():
        if attr != "border":
            setattr(cell, attr, value)


def set_column_widths(ws, widths):
    """Set column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    apply_style(ws["A2"], styles["subheader"])

    instructions = [
        "",
        "PURPOSE",
        "This workbook manages record retention periods and secure disposal processes",
        "as required by ISO 27001:2022 Control A.5.33. Proper retention ensures records",
        "are kept for required periods and disposed of securely when no longer needed.",
        "",
        "KEY RETENTION REGULATIONS (Switzerland)",
        "- Swiss CO Art. 958f: Bookkeeping records - 10 years",
        "- Tax records: 10 years (cantonal variations)",
        "- Employment records: Employment period + 10 years",
        "- GDPR: Purpose-based (storage limitation principle)",
        "",
        "DISPOSAL METHODS BY CLASSIFICATION",
        "- Restricted: Cross-cut shredding P-5, NIST 800-88 Rev. 2 Purge, witnessed destruction",
        "- Confidential: Cross-cut shredding P-4, NIST 800-88 Rev. 2 Clear",
        "- Internal: Standard shredding P-3, secure deletion",
        "- Public: Recycling, standard deletion",
        "",
        "RETENTION PRINCIPLES",
        "1. Minimum Retention: Keep records for legally required period",
        "2. Maximum Retention: Dispose after retention to reduce risk (GDPR)",
        "3. Legal Holds: Override normal retention during litigation",
        "4. Document Everything: Maintain destruction certificates",
        "",
        "HOW TO USE THIS WORKBOOK",
        "1. Define retention periods in Retention_Schedule",
        "2. Map regulations in Regulatory_Mapping",
        "3. Identify records for disposal in Disposal_Queue",
        "4. Apply appropriate methods from Disposal_Method_Matrix",
        "5. Document destruction in Destruction_Verification",
        "6. Track exceptions in Exception_Register",
        "7. Monitor compliance in Compliance_Dashboard",
        "8. Document gaps in Gap_Analysis",
        "9. Obtain approval in Approval_SignOff",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, line in enumerate(instructions, start=4):
        ws[f"A{i}"] = line
        if line in ["PURPOSE", "KEY RETENTION REGULATIONS (Switzerland)",
                    "DISPOSAL METHODS BY CLASSIFICATION", "RETENTION PRINCIPLES",
                    "HOW TO USE THIS WORKBOOK"]:
            ws[f"A{i}"].font = Font(bold=True, size=11)

    set_column_widths(ws, {"A": 100})
    logger.info("Created Instructions sheet")


def create_retention_schedule_sheet(ws, styles):
    """Create the Retention Schedule sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "Retention Schedule"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = "Retention periods for all record categories"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Record Category ID", "Category Name", "Retention Period",
        "Retention Basis", "Basis Detail", "Retention Trigger",
        "Grace Period", "Review Cycle", "Last Review", "Next Review", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Sample data
    sample_data = [
        ["REC-001", "Financial Ledgers", "10 years", "Regulatory",
         "Swiss CO Art. 958f", "Year-End", "90 days", "Annual", "", "", ""],
        ["REC-002", "Personnel Files", "Employment + 10 years", "Regulatory",
         "Swiss employment law", "Employment End", "90 days", "Annual", "", "", "Contains PII"],
        ["REC-003", "Customer Contracts", "Term + 10 years", "Contractual",
         "Standard contract terms", "Contract End", "90 days", "Biennial", "", "", ""],
    ]

    for row_idx, data in enumerate(sample_data, start=5):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Add empty input rows
    for row in range(8, 40):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_basis = DataValidation(
        type="list",
        formula1='"Regulatory,Contractual,Business,Mixed"',
        allow_blank=True
    )
    ws.add_data_validation(dv_basis)
    dv_basis.add("D5:D50")

    dv_trigger = DataValidation(
        type="list",
        formula1='"Creation,Year-End,Contract End,Last Activity,Event-Based"',
        allow_blank=True
    )
    ws.add_data_validation(dv_trigger)
    dv_trigger.add("F5:F50")

    dv_cycle = DataValidation(
        type="list",
        formula1='"Annual,Biennial,Triennial"',
        allow_blank=True
    )
    ws.add_data_validation(dv_cycle)
    dv_cycle.add("H5:H50")

    set_column_widths(ws, {
        "A": 18, "B": 25, "C": 20, "D": 15,
        "E": 30, "F": 18, "G": 12, "H": 12,
        "I": 12, "J": 12, "K": 30
    })
    logger.info("Created Retention_Schedule sheet")


def create_regulatory_mapping_sheet(ws, styles):
    """Create the Regulatory Mapping sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Regulatory Mapping"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Regulations driving record retention requirements"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Regulation ID", "Regulation Name", "Section/Article",
        "Record Types Affected", "Required Retention", "Retention Trigger",
        "Penalty for Non-Compliance", "Last Review", "Reviewer", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Sample data
    sample_data = [
        ["REG-001", "Swiss Code of Obligations", "Art. 958f",
         "Bookkeeping, financial records", "10 years", "Fiscal year end",
         "Administrative fines, audit findings", "", "", ""],
        ["REG-002", "Swiss Federal Tax Act", "Various",
         "Tax records, supporting documents", "10 years", "Tax year end",
         "Tax penalties, audit findings", "", "", ""],
        ["REG-003", "GDPR", "Art. 5(1)(e)",
         "Personal data records", "Purpose-based", "Purpose fulfilled",
         "Up to 4% global revenue", "", "", "Storage limitation principle"],
    ]

    for row_idx, data in enumerate(sample_data, start=5):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Add empty input rows
    for row in range(8, 30):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    set_column_widths(ws, {
        "A": 12, "B": 30, "C": 15, "D": 35,
        "E": 18, "F": 18, "G": 30, "H": 12,
        "I": 20, "J": 30
    })
    logger.info("Created Regulatory_Mapping sheet")


def create_disposal_queue_sheet(ws, styles):
    """Create the Disposal Queue sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "Disposal Queue"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = "Records due or approaching disposal date"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Queue ID", "Record Category", "Retention End Date",
        "Volume - Physical", "Volume - Electronic", "Legal Hold Status",
        "Disposal Priority", "Target Disposal Date", "Assigned To", "Status", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 35):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_hold = DataValidation(
        type="list",
        formula1='"Yes,No,Checking"',
        allow_blank=True
    )
    ws.add_data_validation(dv_hold)
    dv_hold.add("F5:F50")

    dv_priority = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_priority)
    dv_priority.add("G5:G50")

    dv_status = DataValidation(
        type="list",
        formula1='"Pending,In Progress,Complete,On Hold"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("J5:J50")

    set_column_widths(ws, {
        "A": 12, "B": 30, "C": 18, "D": 15,
        "E": 18, "F": 15, "G": 15, "H": 18,
        "I": 20, "J": 15, "K": 30
    })
    logger.info("Created Disposal_Queue sheet")


def create_disposal_method_matrix_sheet(ws, styles):
    """Create the Disposal Method Matrix sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Disposal Method Matrix"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Appropriate disposal methods by classification level"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Classification", "Physical - Paper", "Physical - Media",
        "Electronic - On-Prem", "Electronic - Cloud",
        "Verification Required", "Approved Vendors", "Special Handling"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Pre-populated disposal methods
    methods = [
        ["Restricted", "Cross-cut P-5, witnessed incineration",
         "Degaussing + physical destruction",
         "NIST 800-88 Rev. 2 Purge, cryptographic erasure",
         "Provider purge + verification certificate",
         "Certificate of destruction, witness signature",
         "[List approved vendors]", "Witnessed destruction required"],
        ["Confidential", "Cross-cut shredding P-4",
         "Degaussing or physical destruction",
         "NIST 800-88 Rev. 2 Clear, secure deletion",
         "Provider deletion + confirmation",
         "Certificate of destruction",
         "[List approved vendors]", "Batch processing acceptable"],
        ["Internal", "Standard shredding P-3",
         "Standard deletion or physical destruction",
         "Secure deletion with verification",
         "Standard deletion",
         "Disposal log entry",
         "Internal IT or approved vendor", "Standard process"],
        ["Public", "Recycling acceptable",
         "Standard disposal or recycling",
         "Standard deletion",
         "Standard deletion",
         "Disposal log entry",
         "Any certified recycler", "No special requirements"],
    ]

    for row_idx, method in enumerate(methods, start=5):
        for col_idx, value in enumerate(method, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    set_column_widths(ws, {
        "A": 15, "B": 30, "C": 30, "D": 30,
        "E": 30, "F": 30, "G": 25, "H": 30
    })
    logger.info("Created Disposal_Method_Matrix sheet")


def create_destruction_verification_sheet(ws, styles):
    """Create the Destruction Verification sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Destruction Verification"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Evidence of secure record destruction"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Destruction ID", "Record Category", "Volume",
        "Destruction Date", "Method Used", "Performed By",
        "Witness", "Certificate Reference", "Storage Location", "Verification Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 50):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Pending,Not Verified"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("J5:J60")

    set_column_widths(ws, {
        "A": 15, "B": 25, "C": 15, "D": 15,
        "E": 30, "F": 20, "G": 20, "H": 20,
        "I": 30, "J": 18
    })
    logger.info("Created Destruction_Verification sheet")


def create_exception_register_sheet(ws, styles):
    """Create the Exception Register sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "Exception Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = "Retention extensions and early disposals"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Exception ID", "Exception Type", "Record Category",
        "Original Retention", "New Retention", "Reason",
        "Requested By", "Approved By", "Approval Date", "Expiration", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 30):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_type = DataValidation(
        type="list",
        formula1='"Extension,Early Disposal"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("B5:B40")

    dv_status = DataValidation(
        type="list",
        formula1='"Active,Expired,Cancelled"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("K5:K40")

    set_column_widths(ws, {
        "A": 15, "B": 15, "C": 25, "D": 18,
        "E": 18, "F": 35, "G": 20, "H": 20,
        "I": 15, "J": 18, "K": 12
    })
    logger.info("Created Exception_Register sheet")


def create_compliance_dashboard_sheet(ws, styles):
    """Create the Compliance Dashboard sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Retention Compliance Dashboard"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = "Metrics for monitoring retention schedule adherence"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Metric ID", "Metric Name", "Description", "Target",
        "Current Value", "Trend", "Status", "Owner", "Last Updated"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Sample metrics
    metrics = [
        ["MET-001", "On-Schedule Disposal Rate",
         "% of records disposed within grace period", "95%", "", "", "", "Records Manager", ""],
        ["MET-002", "Overdue Disposals",
         "Number of records past retention + grace", "<5%", "", "", "", "Records Manager", ""],
        ["MET-003", "Legal Hold Compliance",
         "% of holds properly maintained", "100%", "", "", "", "Legal Counsel", ""],
        ["MET-004", "Destruction Certificate Rate",
         "% of disposals with certificates", "100%", "", "", "", "Records Manager", ""],
        ["MET-005", "Retention Schedule Coverage",
         "% of record categories with defined retention", "100%", "", "", "", "Records Manager", ""],
        ["MET-006", "Exception Rate",
         "% of records with retention exceptions", "<5%", "", "", "", "Legal Counsel", ""],
    ]

    for row_idx, metric in enumerate(metrics, start=5):
        for col_idx, value in enumerate(metric, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx in [5, 6, 7]:
                cell.fill = styles["input_cell"]["fill"]

    # Add empty rows
    for row in range(11, 20):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_trend = DataValidation(
        type="list",
        formula1='"Improving,Stable,Declining,New"',
        allow_blank=True
    )
    ws.add_data_validation(dv_trend)
    dv_trend.add("F5:F25")

    dv_status = DataValidation(
        type="list",
        formula1='"On Target,At Risk,Below Target"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("G5:G25")

    set_column_widths(ws, {
        "A": 10, "B": 25, "C": 40, "D": 12,
        "E": 15, "F": 12, "G": 15, "H": 20, "I": 15
    })
    logger.info("Created Compliance_Dashboard sheet")


def create_gap_analysis_sheet(ws, styles):
    """Create the Gap Analysis sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Gap Analysis"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = "Identified retention and disposal issues"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Gap ID", "Gap Category", "Description", "Related Item",
        "Risk Rating", "Remediation Action", "Owner", "Due Date", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 30):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_category = DataValidation(
        type="list",
        formula1='"Retention,Disposal,Verification,Process"',
        allow_blank=True
    )
    ws.add_data_validation(dv_category)
    dv_category.add("B5:B40")

    dv_risk = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_risk)
    dv_risk.add("E5:E40")

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Complete"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("I5:I40")

    set_column_widths(ws, {
        "A": 10, "B": 15, "C": 45, "D": 20,
        "E": 12, "F": 40, "G": 20, "H": 12, "I": 15
    })
    logger.info("Created Gap_Analysis sheet")


def create_evidence_register_sheet(ws, styles):
    """Create the Evidence Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Evidence Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Audit evidence supporting retention and disposal"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Evidence ID", "Description", "Evidence Type",
        "Related Item", "Storage Location", "Collected Date",
        "Collected By", "Verification Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 35):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_type = DataValidation(
        type="list",
        formula1='"Certificate,Log,Approval,Report,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C5:C50")

    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Pending,Expired"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H50")

    set_column_widths(ws, {
        "A": 12, "B": 40, "C": 15, "D": 20,
        "E": 35, "F": 15, "G": 20, "H": 18
    })
    logger.info("Created Evidence_Register sheet")


def create_approval_sheet(ws, styles):
    """Create the Approval Sign-Off sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "Retention and Disposal Schedule Approval"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "Formal approval of retention and disposal assessment"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment metadata
    ws["A4"] = "Assessment Period:"
    ws["B4"] = ""
    ws["A5"] = "Record Categories Covered:"
    ws["B5"] = ""
    ws["A6"] = "Disposal Compliance Rate:"
    ws["B6"] = ""
    ws["A7"] = "Open Gaps:"
    ws["B7"] = ""
    ws["A8"] = "Active Exceptions:"
    ws["B8"] = ""

    for row in range(4, 9):
        ws[f"A{row}"].font = Font(bold=True)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]

    # Approval table
    ws["A11"] = "APPROVAL SIGNATURES"
    ws["A11"].font = Font(bold=True, size=12)

    approval_headers = ["Role", "Name", "Signature", "Date", "Decision", "Comments"]
    for col, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=13, column=col, value=header)
        apply_style(cell, styles["column_header"])

    approvers = [
        "Records Manager",
        "Legal Counsel",
        "Chief Information Security Officer",
        "IT Operations Manager",
        "Compliance Officer",
    ]

    for row_idx, role in enumerate(approvers, start=14):
        ws.cell(row=row_idx, column=1, value=role).border = styles["border"]
        for col in range(2, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected,Pending"',
        allow_blank=True
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add("E14:E20")

    set_column_widths(ws, {"A": 35, "B": 25, "C": 20, "D": 15, "E": 22, "F": 30})
    logger.info("Created Approval_SignOff sheet")


# =============================================================================
# MAIN FUNCTION
# =============================================================================
def main():
    """Main entry point."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS-IMP-A.5.32-33.S3 Retention and Disposal Schedule Generator")
        logger.info("=" * 70)

        wb = Workbook()
        styles = setup_styles()

        # Rename default sheet
        ws_instructions = wb.active
        ws_instructions.title = "Instructions"

        # Create all sheets
        ws_retention = wb.create_sheet("Retention_Schedule")
        ws_regulatory = wb.create_sheet("Regulatory_Mapping")
        ws_queue = wb.create_sheet("Disposal_Queue")
        ws_methods = wb.create_sheet("Disposal_Method_Matrix")
        ws_destruction = wb.create_sheet("Destruction_Verification")
        ws_exceptions = wb.create_sheet("Exception_Register")
        ws_dashboard = wb.create_sheet("Compliance_Dashboard")
        ws_gap = wb.create_sheet("Gap_Analysis")
        ws_evidence = wb.create_sheet("Evidence_Register")
        ws_approval = wb.create_sheet("Approval_SignOff")

        # Populate sheets
        create_instructions_sheet(ws_instructions, styles)
        create_retention_schedule_sheet(ws_retention, styles)
        create_regulatory_mapping_sheet(ws_regulatory, styles)
        create_disposal_queue_sheet(ws_queue, styles)
        create_disposal_method_matrix_sheet(ws_methods, styles)
        create_destruction_verification_sheet(ws_destruction, styles)
        create_exception_register_sheet(ws_exceptions, styles)
        create_compliance_dashboard_sheet(ws_dashboard, styles)
        create_gap_analysis_sheet(ws_gap, styles)
        create_evidence_register_sheet(ws_evidence, styles)
        create_approval_sheet(ws_approval, styles)

        # Save workbook
        wb.save(OUTPUT_FILENAME)
        logger.info("=" * 70)
        logger.info("SUCCESS: Workbook saved as %s", OUTPUT_FILENAME)
        logger.info("=" * 70)
        return 0

    except Exception as e:
        logger.error("FAILED: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.32-33 Information Protection
# =============================================================================
