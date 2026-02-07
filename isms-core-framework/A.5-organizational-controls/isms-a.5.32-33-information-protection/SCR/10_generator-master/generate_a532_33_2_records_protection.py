#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.32-33.S2 - Records Protection Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.33: Protection of Records
"Records should be protected from loss, destruction, falsification, unauthorised
access and unauthorised release in accordance with legislative, regulatory,
contractual and business requirements."

Assessment Domain 2 of 4: Records Protection Assessment

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel workbook for assessing and
documenting the protection controls applied to organisational records.

**Purpose:**
Assess the protection mechanisms ensuring records remain confidential,
maintain integrity, and are available throughout their retention period.

**Generated Workbook Structure:**
1. Instructions - Guidance on records protection assessment
2. Records_Category_Inventory - Record types and classifications
3. Protection_Controls - Controls per record category
4. Integrity_Verification - Integrity testing results
5. Access_Control_Review - Access rights assessment
6. Legal_Hold_Register - Litigation and investigation holds
7. Backup_Verification - Backup and recovery testing
8. Gap_Analysis - Identified gaps and remediation
9. Evidence_Register - Audit evidence tracking
10. Approval_SignOff - Assessment approval

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.32-33.S2"
WORKBOOK_NAME = "Records Protection Assessment"
CONTROL_ID = "A.5.33"
CONTROL_NAME = "Protection of Records"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "compliant": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "warning": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "non_compliant": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    for attr, value in style_dict.items():
        if attr != "border":
            setattr(cell, attr, value)


def set_column_widths(ws, widths):
    """Set column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    apply_style(ws["A2"], styles["subheader"])

    instructions = [
        "",
        "PURPOSE",
        "This workbook assesses the protection controls applied to organisational records",
        "as required by ISO 27001:2022 Control A.5.33. Records must be protected from loss,",
        "destruction, falsification, unauthorised access, and unauthorised release.",
        "",
        "RECORDS PROTECTION PRINCIPLES",
        "- Confidentiality: Access restricted to authorised personnel",
        "- Integrity: Records protected from unauthorised modification",
        "- Availability: Records accessible when needed throughout retention",
        "",
        "RECORD CATEGORIES",
        "- Financial: Ledgers, invoices, tax records (10 years per Swiss CO)",
        "- Personnel: HR files, payroll, benefits (Employment + 10 years)",
        "- Legal: Contracts, litigation files (Duration + 10 years)",
        "- Operational: Project files, meeting minutes (3-7 years)",
        "- Technical: System documentation, configurations (Lifecycle + 3 years)",
        "- Security: Audit logs, incident records (2-7 years)",
        "",
        "HOW TO USE THIS WORKBOOK",
        "1. Complete Records_Category_Inventory with all record types",
        "2. Assess Protection_Controls for each category",
        "3. Test and document Integrity_Verification",
        "4. Review Access_Control_Review for each system",
        "5. Maintain Legal_Hold_Register for active holds",
        "6. Verify Backup_Verification status",
        "7. Document gaps in Gap_Analysis",
        "8. Collect evidence in Evidence_Register",
        "9. Obtain approval in Approval_SignOff",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, line in enumerate(instructions, start=4):
        ws[f"A{i}"] = line
        if line in ["PURPOSE", "RECORDS PROTECTION PRINCIPLES", "RECORD CATEGORIES",
                    "HOW TO USE THIS WORKBOOK"]:
            ws[f"A{i}"].font = Font(bold=True, size=11)

    set_column_widths(ws, {"A": 100})
    logger.info("Created Instructions sheet")


def create_records_category_inventory_sheet(ws, styles):
    """Create the Records Category Inventory sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "Records Category Inventory"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:L2")
    ws["A2"] = "Comprehensive register of all record categories requiring protection"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Record Category ID", "Category Name", "Record Type", "Description",
        "Custodian Department", "Storage Location", "Format",
        "Retention Requirement", "Confidentiality", "Integrity Requirement",
        "Availability Requirement", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Sample data rows
    sample_data = [
        ["REC-001", "Financial Ledgers", "Financial",
         "General ledger, sub-ledgers, journal entries",
         "Finance", "SAP ERP, Archive Server", "Electronic",
         "Swiss CO Art. 958f - 10 years", "Confidential", "Critical",
         "Business Critical", ""],
        ["REC-002", "Personnel Files", "Personnel",
         "Employee HR records, contracts, evaluations",
         "Human Resources", "HRIS, Secure File Share", "Both",
         "Employment + 10 years", "Confidential", "High",
         "Business Critical", "Contains PII"],
        ["REC-003", "Security Audit Logs", "Security",
         "System access logs, security events",
         "IT Security", "SIEM, Log Archive", "Electronic",
         "7 years", "Confidential", "Critical",
         "Business Critical", "Integrity critical for audit"],
    ]

    for row_idx, data in enumerate(sample_data, start=5):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Add empty input rows
    for row in range(8, 35):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_type = DataValidation(
        type="list",
        formula1='"Financial,Personnel,Legal,Operational,Technical,Security,Regulatory"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C5:C50")

    dv_format = DataValidation(
        type="list",
        formula1='"Physical,Electronic,Both"',
        allow_blank=True
    )
    ws.add_data_validation(dv_format)
    dv_format.add("G5:G50")

    dv_conf = DataValidation(
        type="list",
        formula1='"Restricted,Confidential,Internal,Public"',
        allow_blank=True
    )
    ws.add_data_validation(dv_conf)
    dv_conf.add("I5:I50")

    dv_integrity = DataValidation(
        type="list",
        formula1='"Critical,High,Standard"',
        allow_blank=True
    )
    ws.add_data_validation(dv_integrity)
    dv_integrity.add("J5:J50")

    dv_avail = DataValidation(
        type="list",
        formula1='"Mission Critical,Business Critical,Standard"',
        allow_blank=True
    )
    ws.add_data_validation(dv_avail)
    dv_avail.add("K5:K50")

    set_column_widths(ws, {
        "A": 18, "B": 25, "C": 12, "D": 40,
        "E": 18, "F": 30, "G": 12, "H": 30,
        "I": 15, "J": 15, "K": 18, "L": 30
    })
    logger.info("Created Records_Category_Inventory sheet")


def create_protection_controls_sheet(ws, styles):
    """Create the Protection Controls sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Protection Controls Assessment"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Confidentiality, integrity, and availability controls per record category"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Record Category ID", "Category Name", "Confidentiality Controls",
        "Integrity Controls", "Availability Controls", "Physical Controls",
        "Control Effectiveness", "Gap Description", "Remediation Needed", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 35):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_eff = DataValidation(
        type="list",
        formula1='"Effective,Partial,Ineffective"',
        allow_blank=True
    )
    ws.add_data_validation(dv_eff)
    dv_eff.add("G5:G50")

    dv_status = DataValidation(
        type="list",
        formula1='"Complete,In Progress,Not Started"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("J5:J50")

    set_column_widths(ws, {
        "A": 18, "B": 20, "C": 35, "D": 35,
        "E": 35, "F": 25, "G": 18, "H": 35,
        "I": 35, "J": 15
    })
    logger.info("Created Protection_Controls sheet")


def create_integrity_verification_sheet(ws, styles):
    """Create the Integrity Verification sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Integrity Verification"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Testing and verification of record integrity mechanisms"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Test ID", "Record Category", "Integrity Mechanism", "Test Date",
        "Test Performed", "Expected Result", "Actual Result",
        "Status", "Issues", "Remediation"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 30):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_mechanism = DataValidation(
        type="list",
        formula1='"Checksum,Digital Signature,WORM,Audit Log,Database Constraints,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_mechanism)
    dv_mechanism.add("C5:C40")

    dv_status = DataValidation(
        type="list",
        formula1='"Pass,Fail,Partial,Not Tested"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H40")

    set_column_widths(ws, {
        "A": 10, "B": 25, "C": 20, "D": 12,
        "E": 35, "F": 25, "G": 25, "H": 12,
        "I": 30, "J": 30
    })
    logger.info("Created Integrity_Verification sheet")


def create_access_control_review_sheet(ws, styles):
    """Create the Access Control Review sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "Access Control Review"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = "Assessment of access controls for record storage systems"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "System Name", "Record Categories", "Access Control Type",
        "User Count", "Privileged Users", "Last Access Review",
        "Access Logging", "Log Retention", "Issues", "Remediation", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 25):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_type = DataValidation(
        type="list",
        formula1='"RBAC,DAC,MAC,Mixed"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C5:C30")

    dv_logging = DataValidation(
        type="list",
        formula1='"Yes,No,Partial"',
        allow_blank=True
    )
    ws.add_data_validation(dv_logging)
    dv_logging.add("G5:G30")

    dv_status = DataValidation(
        type="list",
        formula1='"Compliant,Non-Compliant,Partial"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("K5:K30")

    set_column_widths(ws, {
        "A": 25, "B": 30, "C": 15, "D": 12,
        "E": 15, "F": 18, "G": 15, "H": 15,
        "I": 30, "J": 30, "K": 15
    })
    logger.info("Created Access_Control_Review sheet")


def create_legal_hold_register_sheet(ws, styles):
    """Create the Legal Hold Register sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Legal Hold Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Litigation and investigation holds affecting record retention"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Hold ID", "Matter Name", "Legal Counsel", "Effective Date",
        "Record Categories", "Custodians Notified", "Notification Date",
        "Release Date", "Status", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 25):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_status = DataValidation(
        type="list",
        formula1='"Active,Released,Pending"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("I5:I30")

    set_column_widths(ws, {
        "A": 15, "B": 30, "C": 25, "D": 15,
        "E": 35, "F": 30, "G": 15, "H": 15,
        "I": 12, "J": 35
    })
    logger.info("Created Legal_Hold_Register sheet")


def create_backup_verification_sheet(ws, styles):
    """Create the Backup Verification sheet."""
    ws.merge_cells("A1:M1")
    ws["A1"] = "Backup Verification"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:M2")
    ws["A2"] = "Backup and recovery verification for records availability"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Record Category", "Backup System", "Backup Frequency", "Backup Location",
        "Last Backup Date", "Last Verification", "Last Recovery Test",
        "RTO Target", "RTO Achieved", "RPO Target", "RPO Achieved",
        "Status", "Issues"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 30):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_freq = DataValidation(
        type="list",
        formula1='"Real-time,Hourly,Daily,Weekly,Monthly"',
        allow_blank=True
    )
    ws.add_data_validation(dv_freq)
    dv_freq.add("C5:C40")

    dv_status = DataValidation(
        type="list",
        formula1='"Compliant,Non-Compliant"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("L5:L40")

    set_column_widths(ws, {
        "A": 20, "B": 18, "C": 15, "D": 25,
        "E": 15, "F": 15, "G": 18, "H": 12,
        "I": 12, "J": 12, "K": 12, "L": 15, "M": 30
    })
    logger.info("Created Backup_Verification sheet")


def create_gap_analysis_sheet(ws, styles):
    """Create the Gap Analysis sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Gap Analysis"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Identified protection gaps with remediation tracking"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Gap ID", "Gap Category", "Description", "Related Record Category",
        "Risk Rating", "Remediation Action", "Owner",
        "Due Date", "Status", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 35):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_category = DataValidation(
        type="list",
        formula1='"Confidentiality,Integrity,Availability,Process"',
        allow_blank=True
    )
    ws.add_data_validation(dv_category)
    dv_category.add("B5:B50")

    dv_risk = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_risk)
    dv_risk.add("E5:E50")

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Complete,Accepted"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("I5:I50")

    set_column_widths(ws, {
        "A": 10, "B": 15, "C": 45, "D": 20,
        "E": 12, "F": 40, "G": 20, "H": 12,
        "I": 15, "J": 30
    })
    logger.info("Created Gap_Analysis sheet")


def create_evidence_register_sheet(ws, styles):
    """Create the Evidence Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Evidence Register"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Audit evidence supporting records protection assessment"
    apply_style(ws["A2"], styles["subheader"])

    headers = [
        "Evidence ID", "Description", "Evidence Type",
        "Related Item", "Storage Location", "Collected Date",
        "Collected By", "Verification Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_style(cell, styles["column_header"])

    # Add empty input rows
    for row in range(5, 35):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_type = DataValidation(
        type="list",
        formula1='"Document,Screenshot,Report,Log,Configuration,Other"',
        allow_blank=True
    )
    ws.add_data_validation(dv_type)
    dv_type.add("C5:C50")

    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Pending Review,Not Verified,Expired"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H50")

    set_column_widths(ws, {
        "A": 12, "B": 40, "C": 18, "D": 20,
        "E": 35, "F": 15, "G": 20, "H": 18
    })
    logger.info("Created Evidence_Register sheet")


def create_approval_sheet(ws, styles):
    """Create the Approval Sign-Off sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "Records Protection Assessment Approval"
    apply_style(ws["A1"], styles["header"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "Formal approval of records protection assessment"
    apply_style(ws["A2"], styles["subheader"])

    # Assessment metadata
    ws["A4"] = "Assessment Period:"
    ws["B4"] = ""
    ws["A5"] = "Records Categories Assessed:"
    ws["B5"] = ""
    ws["A6"] = "Protection Effectiveness:"
    ws["B6"] = ""
    ws["A7"] = "Open Gaps:"
    ws["B7"] = ""
    ws["A8"] = "Active Legal Holds:"
    ws["B8"] = ""

    for row in range(4, 9):
        ws[f"A{row}"].font = Font(bold=True)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]

    # Approval table
    ws["A11"] = "APPROVAL SIGNATURES"
    ws["A11"].font = Font(bold=True, size=12)

    approval_headers = ["Role", "Name", "Signature", "Date", "Decision", "Comments"]
    for col, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=13, column=col, value=header)
        apply_style(cell, styles["column_header"])

    approvers = [
        "Records Manager",
        "Chief Information Security Officer",
        "Legal Counsel",
        "IT Operations Manager",
        "Internal Audit Representative",
    ]

    for row_idx, role in enumerate(approvers, start=14):
        ws.cell(row=row_idx, column=1, value=role).border = styles["border"]
        for col in range(2, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            cell.fill = styles["input_cell"]["fill"]

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected,Pending"',
        allow_blank=True
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add("E14:E20")

    set_column_widths(ws, {"A": 35, "B": 25, "C": 20, "D": 15, "E": 22, "F": 30})
    logger.info("Created Approval_SignOff sheet")


# =============================================================================
# MAIN FUNCTION
# =============================================================================
def main():
    """Main entry point."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS-IMP-A.5.32-33.S2 Records Protection Assessment Generator")
        logger.info("=" * 70)

        wb = Workbook()
        styles = setup_styles()

        # Rename default sheet
        ws_instructions = wb.active
        ws_instructions.title = "Instructions"

        # Create all sheets
        ws_inventory = wb.create_sheet("Records_Category_Inventory")
        ws_controls = wb.create_sheet("Protection_Controls")
        ws_integrity = wb.create_sheet("Integrity_Verification")
        ws_access = wb.create_sheet("Access_Control_Review")
        ws_hold = wb.create_sheet("Legal_Hold_Register")
        ws_backup = wb.create_sheet("Backup_Verification")
        ws_gap = wb.create_sheet("Gap_Analysis")
        ws_evidence = wb.create_sheet("Evidence_Register")
        ws_approval = wb.create_sheet("Approval_SignOff")

        # Populate sheets
        create_instructions_sheet(ws_instructions, styles)
        create_records_category_inventory_sheet(ws_inventory, styles)
        create_protection_controls_sheet(ws_controls, styles)
        create_integrity_verification_sheet(ws_integrity, styles)
        create_access_control_review_sheet(ws_access, styles)
        create_legal_hold_register_sheet(ws_hold, styles)
        create_backup_verification_sheet(ws_backup, styles)
        create_gap_analysis_sheet(ws_gap, styles)
        create_evidence_register_sheet(ws_evidence, styles)
        create_approval_sheet(ws_approval, styles)

        # Save workbook
        wb.save(OUTPUT_FILENAME)
        logger.info("=" * 70)
        logger.info("SUCCESS: Workbook saved as %s", OUTPUT_FILENAME)
        logger.info("=" * 70)
        return 0

    except Exception as e:
        logger.error("FAILED: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.32-33 Information Protection
# =============================================================================
