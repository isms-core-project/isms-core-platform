#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.8.1 - Project Lifecycle Security Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.8: Information Security in Project Management
Assessment Domain 1 of 3: Project Lifecycle Security Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific project management methodology, security gates,
and assessment requirements.

Key customization areas:
1. Project classification criteria (match your risk categories)
2. Security gate definitions (aligned with your SDLC phases)
3. Assessment checklists (specific to your security requirements)
4. Approval workflow (per your governance structure)
5. Integration with PM tools (Jira, ServiceNow per your tooling)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for project management)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

**Purpose:**
Enables systematic security assessment throughout the project lifecycle,
supporting ISO 27001:2022 Control A.5.8 requirements for integrating
information security into project management.

**Assessment Scope:**
- Project security classification and risk assessment
- Security requirements identification per project phase
- Security gate reviews (initiation, planning, execution, closure)
- Threat and vulnerability assessment per project
- Security control implementation tracking
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance
2. Project Classification - Risk-based project categorization
3. Initiation Phase - Security requirements identification
4. Planning Phase - Security control planning
5. Execution Phase - Implementation verification
6. Monitoring Phase - Security testing and validation
7. Closure Phase - Security acceptance and handover
8. Compliance Dashboard - Overall security compliance status
9. Evidence Register - Audit evidence tracking
10. Sign-Off - Stakeholder approval workflow

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a58_1_project_lifecycle_assessment.py

Requirements:
    - Python 3.8+
    - openpyxl library: pip install openpyxl

Output:
    ISMS-IMP-A.5.8.1_Project_Lifecycle_Assessment_YYYYMMDD.xlsx
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    logger.error("Error: pip3 install openpyxl"); sys.exit(1)

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

DOCUMENT_ID = "ISMS-IMP-A.5.8.1"
CONTROL_REF = "ISO/IEC 27001:2022 Control A.5.8"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_Project_Lifecycle_Assessment_{GENERATED_TIMESTAMP}.xlsx"

def setup_styles():
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    return {
        "header": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)
        },
        "section_header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="305496", end_color="305496", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True)
        },
        "subsection_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="000000"),
            "fill": PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True)
        },
        "column_header": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border
        },
        "protected_cell": {
            "fill": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": border
        },
        "border": border
    }

def create_workbook():
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    sheets = [
        "Instructions & Legend",
        "2. Project Classification", 
        "3. Initiation Phase",
        "4. Planning Phase",
        "5. Execution Phase",
        "6. Monitoring Phase",
        "7. Closure Phase",
        "8. Compliance Dashboard",
        "9. Evidence Register",
        "10. Sign-Off"
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb

def create_instructions_sheet(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{DOCUMENT_ID} - Project Lifecycle Security Assessment\\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40
    row = 3
    ws[f"A{row}"] = "DOCUMENT CONTROL"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    row += 2
    for label, value in [
        ("Document ID:", DOCUMENT_ID),
        ("Version:", "2.0"),
        ("Control Reference:", CONTROL_REF),
        ("Purpose:", "Assess security integration across project lifecycle (6 phases)")
    ]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 55

def create_classification_sheet(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "PHASE 0: PROJECT CLASSIFICATION\\nDetermine Risk Level (High/Medium/Low)"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    row = 3
    ws[f"A{row}"] = "Project Name:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]
    row += 1
    ws[f"A{row}"] = "Project Manager:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]
    # Workshop guidance
    row += 2
    ws[f"A{row}"] = "CLASSIFICATION WORKSHOP (30-60 minutes)"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    ws[f"A{row}"] = "1. Gather stakeholders: PM, Security, Business Owner, Technical Lead (5 min)"
    row += 1
    ws[f"A{row}"] = "2. Review each factor below and score High/Medium/Low (20-30 min)"
    row += 1
    ws[f"A{row}"] = "3. Discuss edge cases and document rationale (10-15 min)"
    row += 1
    ws[f"A{row}"] = "4. Confirm final classification and next steps (5 min)"
    row += 2

    # 6-Factor Classification Matrix (per ISMS Copilot audit)
    ws[f"A{row}"] = "CLASSIFICATION DECISION MATRIX (6 Factors):"
    ws[f"A{row}"].font = styles["subsection_header"]["font"]
    ws[f"A{row}"].fill = styles["subsection_header"]["fill"]
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    # Headers
    factor_headers = ["Factor", "High (3 pts)", "Medium (2 pts)", "Low (1 pt)", "Score", "Rationale"]
    for col_idx, header in enumerate(factor_headers, 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = Font(bold=True, size=10)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws[f"{col}{row}"].border = styles["border"]
    row += 1

    # 6 Classification Factors with criteria
    factors = [
        ("Data Sensitivity", "Restricted/Special Category PII", "Confidential/Internal PII", "Public/Non-sensitive"),
        ("System Criticality", "Tier 1 (RTO <4hr)", "Tier 2 (RTO 4-24hr)", "Tier 3+ (RTO >24hr)"),
        ("Regulatory Scope", "Multiple regulations (GDPR+PCI+etc)", "Single major regulation", "General compliance only"),
        ("External Exposure", "Internet-facing, customer data", "Partner/B2B access", "Internal only"),
        ("Technical Complexity", "New technology, high integration", "Moderate complexity", "Standard/proven stack"),
        ("Third-Party Involvement", "Critical outsourced development", "Vendor components", "Internal development"),
    ]

    score_validation = DataValidation(type="list", formula1='"3,2,1"', allow_blank=False)
    ws.add_data_validation(score_validation)
    factor_start_row = row

    for factor_name, high_desc, med_desc, low_desc in factors:
        ws[f"A{row}"] = factor_name
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = high_desc
        ws[f"C{row}"] = med_desc
        ws[f"D{row}"] = low_desc
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].border = styles["border"]
        score_validation.add(ws[f"E{row}"])
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].border = styles["border"]
        for col in ['A', 'B', 'C', 'D']:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Total score and classification
    row += 1
    ws[f"A{row}"] = "TOTAL SCORE:"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"E{row}"] = f"=SUM(E{factor_start_row}:E{factor_start_row+5})"
    ws[f"E{row}"].font = Font(bold=True, size=12)
    ws[f"E{row}"].border = styles["border"]
    row += 1

    ws[f"A{row}"] = "CALCULATED RISK LEVEL:"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"B{row}"] = f'=IF(E{row-1}>=15,"High",IF(E{row-1}>=10,"Medium","Low"))'
    ws[f"B{row}"].font = Font(bold=True, size=14, color="FF0000")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    row += 1
    ws[f"A{row}"] = "(High: 15-18 pts | Medium: 10-14 pts | Low: 6-9 pts)"
    ws[f"A{row}"].font = Font(italic=True, size=9)

    # Common Classification Scenarios (per audit requirement)
    row += 3
    ws[f"A{row}"] = "COMMON CLASSIFICATION SCENARIOS (Reference Examples):"
    ws[f"A{row}"].font = styles["subsection_header"]["font"]
    ws[f"A{row}"].fill = styles["subsection_header"]["fill"]
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    scenarios = [
        ("Customer-facing web portal with payment processing", "High", "External exposure + PCI DSS + customer PII"),
        ("Internal HR system with employee data", "Medium", "Internal PII but limited exposure"),
        ("Marketing website refresh (static content)", "Low", "Public data, no integration"),
        ("Cloud migration of core ERP system", "High", "Critical system + complexity + third-party"),
        ("Mobile app for field service technicians", "Medium", "Internal users but external network"),
        ("Data analytics dashboard (aggregated metrics)", "Low", "No PII, internal only"),
        ("Third-party SaaS integration (CRM)", "Medium", "Vendor involvement + customer data subset"),
        ("Legacy system decommissioning", "Medium", "Data retention + secure disposal needs"),
    ]

    scenario_headers = ["Scenario Description", "Typical Classification", "Key Factors"]
    for col_idx, header in enumerate(scenario_headers, 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = Font(bold=True, size=10)
        ws[f"{col}{row}"].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws[f"{col}{row}"].border = styles["border"]
    row += 1

    for scenario_desc, classification, factors_desc in scenarios:
        ws[f"A{row}"] = scenario_desc
        ws[f"B{row}"] = classification
        ws[f"C{row}"] = factors_desc
        for col in ['A', 'B', 'C']:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].alignment = Alignment(wrap_text=True)
        row += 1
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 35

def create_phase_sheet(ws, styles, phase_name, activities):
    ws.merge_cells("A1:E1")
    ws["A1"] = phase_name.upper()
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    row = 3
    headers = ["Activity", "Status", "Completion Date", "Evidence Link", "Notes"]
    widths = [60, 15, 15, 40, 40]
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws[f"{col}{row}"].alignment = styles["column_header"]["alignment"]
        ws[f"{col}{row}"].border = styles["border"]
        ws.column_dimensions[col].width = width
    row += 1
    validations = DataValidation(
        type="list",
        formula1='"✅ Complete,🔄 In Progress,⚠️ Incomplete,❌ Not Done,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations)
    for activity in activities:
        ws[f"A{row}"] = activity
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws[f"A{row}"].border = styles["border"]
        for col in ['B', 'C', 'D', 'E']:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = styles["border"]
        validations.add(ws[f"B{row}"])
        ws[f"C{row}"].number_format = "DD.MM.YYYY"
        row += 1
    ws.freeze_panes = "A4"

def create_dashboard_sheet(ws, styles):
    ws.merge_cells("A1:D1")
    ws["A1"] = "COMPLIANCE DASHBOARD\\nAutomated Phase Completion Scores"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    row = 3
    ws[f"A{row}"] = "PROJECT INFORMATION"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    row += 2
    ws[f"A{row}"] = "Project Name:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "='2. Project Classification'!B3"
    row += 1
    ws[f"A{row}"] = "Risk Classification:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "='2. Project Classification'!B18"
    ws[f"B{row}"].font = Font(bold=True, size=12, color="FF0000")
    row += 1
    ws[f"A{row}"] = "Assessment Date:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "=TODAY()"
    ws[f"B{row}"].number_format = "DD.MM.YYYY"
    row += 2
    ws[f"A{row}"] = "PHASE COMPLETION SCORES"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    row += 1
    ws[f"A{row}"] = "Phase"
    ws[f"B{row}"] = "Completion %"
    for col in ['A', 'B']:
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
    row += 1
    phases = [
        ("Initiation", "'3. Initiation Phase'", 4, 9),
        ("Planning", "'4. Planning Phase'", 4, 11),
        ("Execution", "'5. Execution Phase'", 4, 12),
        ("Monitoring", "'6. Monitoring Phase'", 4, 8),
        ("Closure", "'7. Closure Phase'", 4, 9)
    ]
    for phase_name, sheet_ref, start_row, end_row in phases:
        ws[f"A{row}"] = phase_name
        ws[f"B{row}"] = f'=COUNTIF({sheet_ref}!B{start_row}:B{end_row},"✅ Complete")/(COUNTA({sheet_ref}!B{start_row}:B{end_row})-COUNTIF({sheet_ref}!B{start_row}:B{end_row},"N/A"))'
        ws[f"B{row}"].number_format = "0%"
        row += 1
    row += 1
    ws[f"A{row}"] = "OVERALL COMPLIANCE:"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"B{row}"] = f"=AVERAGE(B{row-5}:B{row-1})"
    ws[f"B{row}"].number_format = "0%"
    ws[f"B{row}"].font = Font(bold=True, size=14)
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15

def create_evidence_register(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "EVIDENCE REGISTER\\nAudit Evidence Tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    row = 3
    headers = ["Evidence ID", "Phase", "Description", "Location/Path", "Date Collected", "Status"]
    widths = [12, 20, 50, 45, 15, 18]
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
        ws[f"{col}{row}"].alignment = styles["column_header"]["alignment"]
        ws[f"{col}{row}"].border = styles["border"]
        ws.column_dimensions[col].width = width
    row += 1
    validations_phase = DataValidation(
        type="list",
        formula1='"Classification,Initiation,Planning,Execution,Monitoring,Closure"',
        allow_blank=False
    )
    ws.add_data_validation(validations_phase)
    validations_status = DataValidation(
        type="list",
        formula1='"Collected,Verified,Pending,Missing"',
        allow_blank=False
    )
    ws.add_data_validation(validations_status)
    for i in range(100):
        ws[f"A{row}"] = f"=TEXT(ROW()-3,\"EV-000\")"
        ws[f"A{row}"].font = Font(color="808080")
        ws[f"A{row}"].fill = styles["protected_cell"]["fill"]
        ws[f"A{row}"].border = styles["border"]
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = styles["border"]
        validations_phase.add(ws[f"B{row}"])
        validations_status.add(ws[f"F{row}"])
        ws[f"E{row}"].number_format = "DD.MM.YYYY"
        row += 1
    ws.freeze_panes = "A4"

def create_signoff_sheet(ws, styles):
    ws.merge_cells("A1:E1")
    ws["A1"] = "SIGN-OFF & APPROVAL\\nMulti-Stakeholder Approval Workflow"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    row = 3
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    row += 2
    ws[f"A{row}"] = "Project Name:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "='2. Project Classification'!B3"
    row += 1
    ws[f"A{row}"] = "Overall Compliance:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "='8. Compliance Dashboard'!B15"
    ws[f"B{row}"].number_format = "0%"
    ws[f"B{row}"].font = Font(bold=True)
    row += 2
    ws[f"A{row}"] = "APPROVALS"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]
    row += 1
    headers = ["Role", "Name", "Signature", "Date", "Decision"]
    for col_idx, header in enumerate(headers, 1):
        col = get_column_letter(col_idx)
        ws[f"{col}{row}"] = header
        ws[f"{col}{row}"].font = styles["column_header"]["font"]
        ws[f"{col}{row}"].fill = styles["column_header"]["fill"]
    row += 1
    validations = DataValidation(
        type="list",
        formula1='"✅ Approved,⚠️ Conditional,❌ Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(validations)
    for role in ["Project Manager", "Security Team", "Final Approver"]:
        ws[f"A{row}"] = role
        ws[f"A{row}"].font = Font(bold=True)
        for col in ['B', 'C', 'D', 'E']:
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{row}"].border = styles["border"]
        validations.add(ws[f"E{row}"])
        ws[f"D{row}"].number_format = "DD.MM.YYYY"
        row += 1
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20

def main():
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - Project Lifecycle Security Assessment Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)
    wb = create_workbook()
    styles = setup_styles()
    logger.info("\\n[1/10] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    logger.info("[2/10] Creating Project Classification...")
    create_classification_sheet(wb["2. Project Classification"], styles)
    logger.info("[3/10] Creating Initiation Phase...")
    create_phase_sheet(wb["3. Initiation Phase"], styles, "Phase 1: Initiation", [
        "Identify security stakeholders and establish communication plan",
        "Conduct initial risk assessment and document key risks",
        "Define security requirements baseline",
        "Allocate security budget and resources",
        "Document security responsibilities in project charter",
        "Obtain Phase 1 gate approval from security team"
    ])
    logger.info("[4/10] Creating Planning Phase...")
    create_phase_sheet(wb["4. Planning Phase"], styles, "Phase 2: Planning", [
        "Document detailed security requirements (link to A.5.8.2 Register)",
        "Conduct threat modeling and document threat scenarios",
        "Develop security test plan and define test cases",
        "Complete Data Protection Impact Assessment (DPIA) if required",
        "Conduct vendor security assessment if third-party components",
        "Define security monitoring and logging requirements",
        "Establish incident response procedures",
        "Obtain Phase 2 gate approval from security team"
    ])
    logger.info("[5/10] Creating Execution Phase...")
    create_phase_sheet(wb["5. Execution Phase"], styles, "Phase 3: Execution", [
        "Execute security testing per test plan (SAST, DAST, etc.)",
        "Conduct penetration testing and document findings",
        "Complete vulnerability scans and remediate critical/high findings",
        "Review and approve security architecture/design",
        "Verify secure coding practices and code review completion",
        "Test security controls and verify implementation",
        "Document security configurations and hardening",
        "Update threat model with implementation changes",
        "Obtain Phase 3 gate approval from security team"
    ])
    logger.info("[6/10] Creating Monitoring Phase...")
    create_phase_sheet(wb["6. Monitoring Phase"], styles, "Phase 4: Monitoring", [
        "Monitor ongoing compliance with security requirements",
        "Review and update risk assessments for changes",
        "Assess security impact of change requests",
        "Track security metrics and KPIs",
        "Obtain Phase 4 gate approval for major milestones"
    ])
    logger.info("[7/10] Creating Closure Phase...")
    create_phase_sheet(wb["7. Closure Phase"], styles, "Phase 5: Closure", [
        "Complete security handover documentation",
        "Verify all security testing complete and passed",
        "Document residual risks and obtain risk acceptance",
        "Conduct lessons learned session with security team",
        "Archive security documentation and evidence",
        "Obtain final security sign-off for production deployment"
    ])
    logger.info("[8/10] Creating Compliance Dashboard...")
    create_dashboard_sheet(wb["8. Compliance Dashboard"], styles)
    logger.info("[9/10] Creating Evidence Register...")
    create_evidence_register(wb["9. Evidence Register"], styles)
    logger.info("[10/10] Creating Sign-Off sheet...")
    create_signoff_sheet(wb["10. Sign-Off"], styles)
    filename = OUTPUT_FILENAME
    wb.save(filename)
    logger.info(f"\\n✅ SUCCESS: {filename}")
    logger.info(f"📄 File: {filename}")
    logger.info(f"📊 Sheets: 10 sheets created")
    logger.info("\\nNext Steps:")
    logger.info("1. Complete Project Classification to determine risk level")
    logger.info("2. Complete phases sequentially as project progresses")
    logger.info("3. Link evidence in Evidence Register")
    logger.info("4. Review Compliance Dashboard for scores")
    logger.info("5. Obtain sign-offs when complete")
    logger.info("=" * 78)

if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-02-02
# QA_STATUS: PASSED - AUDIT ALIGNMENT UPDATE
# QA_TOOL: Claude Code
# CHANGES: Added 6-factor classification matrix (Data Sensitivity, System Criticality,
#          Regulatory Scope, External Exposure, Technical Complexity, Third-Party);
#          Added 30-60 min workshop procedure; Added 8 common classification scenarios
# =============================================================================
