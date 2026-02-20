#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.23.S1 - Cloud Service Inventory & Classification Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.23: Information Security for Use of Cloud Services
Assessment Domain 1 of 5: Cloud Service Inventory & Classification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific cloud service landscape, regulatory requirements,
and classification criteria.

Key customization areas:
1. Cloud provider inventory (AWS, Azure, GCP per your deployment)
2. Service classification criteria (aligned with your data classification)
3. Regulatory mappings (DORA, NIS2, AI Act per your jurisdictions)
4. Risk rating criteria (based on your risk framework)
5. Data residency requirements (per your compliance obligations)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for cloud services)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

**Purpose:**
Enables systematic inventory and classification of cloud services supporting
ISO 27001:2022 Control A.5.23 requirements and regulatory compliance (DORA,
NIS2, AI Act, CLOUD Act).

**Assessment Scope:**
- Cloud service inventory (IaaS, PaaS, SaaS across providers)
- Service classification (criticality, data sensitivity)
- Regulatory applicability mapping
- Data residency and sovereignty tracking
- Concentration risk assessment
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance
2. Cloud Service Inventory - Complete service catalog
3. Classification Matrix - Criticality and sensitivity
4. Regulatory Mapping - DORA, NIS2, AI Act applicability
5. Data Residency - Sovereignty and location tracking
6. Concentration Risk - Provider dependency analysis
7. Gap Analysis - Non-compliant services
8. Evidence Register - Audit evidence tracking
9. Approval & Sign-Off - Stakeholder approval

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a523_1_inventory.py

Requirements:
    - Python 3.8+
    - openpyxl library: pip install openpyxl

Output:
    ISMS-IMP-A.5.23.S1_Inventory_DD_YYYYMMDD.xlsx
"""

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.23.S1"
WORKBOOK_NAME = "Cloud Service Inventory & Classification"
CONTROL_ID = "A.5.19-23"
CONTROL_NAME = "Information Security for Use of Cloud Services"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# REGULATORY DROPDOWN DEFINITIONS (DORA, NIS2, AI Act, CLOUD Act)
# ============================================================================

PROVIDER_HQ_JURISDICTION = [
    "Switzerland", "EU/EEA", "United Kingdom", "United States",
    "Other Adequate Country", "Non-Adequate Country"
]

CLOUD_ACT_EXPOSURE = [
    "No Exposure", "Potential Exposure (US HQ)", "Mitigated (EU Data Boundary)",
    "Mitigated (Encryption + Key Control)", "Accepted Risk (Documented)", "Under Assessment"
]

AI_RISK_CLASSIFICATION = [
    "Not Applicable", "Minimal Risk", "Limited Risk (Transparency)",
    "High Risk", "Unacceptable Risk (Prohibited)"
]

CONCENTRATION_RISK_LEVEL = [
    "Low (Multiple alternatives)", "Medium (Limited alternatives)",
    "High (Few alternatives)", "Critical (Single provider dependency)"
]

EU_DATA_BOUNDARY_OPTIONS = ["Yes", "No", "Partial", "Unknown"]
YES_NO_NOT_DETERMINED = ["Yes", "No", "Not Determined"]
YES_NO_PARTIAL = ["Yes", "No", "Partial"]

# Exit Strategy Framework (POL-S5 Section 8.1.1)
EXIT_STRATEGY_TYPE = [
    "Cloud-to-Cloud", 
    "Hybrid", 
    "On-Premises", 
    "Not Yet Determined"
]

ALTERNATIVE_IDENTIFIED = [
    "Cloud Provider (AWS)", 
    "Cloud Provider (Azure)", 
    "Cloud Provider (GCP)",
    "Cloud Provider (OVHcloud)",
    "Cloud Provider (Alibaba)",
    "Cloud Provider (Other)",
    "Hybrid (Cloud+OnPrem)", 
    "On-Prem Only", 
    "Multiple Options", 
    "Not Yet Assessed"
]

EXPORT_FORMAT = [
    "Standard (CSV/JSON)", 
    "Proprietary", 
    "API Only", 
    "None"
]

EXPORT_TESTED = ["Yes", "No", "Partial"]

MIGRATION_COMPLEXITY = [
    "Cloud-to-Cloud (Low)",
    "Cloud-to-Cloud (Medium)",
    "Cloud-to-Cloud (High)",
    "Hybrid (Medium)",
    "Hybrid (High)",
    "On-Prem (High)",
    "On-Prem (Very High)"
]

LOCK_IN_RISK = ["Low", "Medium", "High", "Critical"]

HYBRID_WORKLOAD_SEGMENTATION = ["Documented", "In Progress", "Not Applicable"]

HYBRID_DATA_SYNC_LATENCY = [
    "Excellent (<100ms)", 
    "Acceptable (100-500ms)", 
    "Concern (>500ms)", 
    "Not Tested", 
    "N/A"
]

ONPREM_TCO_ANALYSIS = [
    "Yes (Favorable)", 
    "Yes (Unfavorable)", 
    "In Progress", 
    "Not Started", 
    "N/A"
]

ONPREM_INFRASTRUCTURE = [
    "Yes (Sufficient)", 
    "Partial (Upgrade Needed)", 
    "No (Full Build)", 
    "Not Assessed", 
    "N/A"
]

ONPREM_STAFFING_PLAN = ["Yes", "In Progress", "Not Started", "N/A"]

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠ Warning sign
SHIELD = '\u2660'     # ♠ Shield (BMP)
LOCK = '\u2302'       # ⌂ Lock (BMP)
CLOUD = '\u2601'      # ☁ Cloud
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# SECTION 1: WORKBOOK CREATION
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    sheets = [
        "Instructions & Legend",
        "2. SaaS Services",
        "3. IaaS PaaS Services",
        "4. Cloud Security Services",
        "5. Cloud Storage Services",
        "6. Data Classification",
        "7. Service Criticality",
        "8. Summary Dashboard",
        "9. Evidence Register",
        "10. Approval Sign-Of",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================

def setup_styles() -> dict:
    """Define cell styles - returns dict of style definitions."""
    return {
        "header": {
            "font_params": {"name": "Calibri", "size": 14, "bold": True, "color": "FFFFFF"},
            "fill_params": {"start_color": "003366", "end_color": "003366", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "subheader": {
            "font_params": {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"},
            "fill_params": {"start_color": "4472C4", "end_color": "4472C4", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "column_header": {
            "font_params": {"name": "Calibri", "size": 10, "bold": True},
            "fill_params": {"start_color": "D9D9D9", "end_color": "D9D9D9", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "input_cell": {
            "fill_params": {"start_color": "FFFFCC", "end_color": "FFFFCC", "fill_type": "solid"},
            "alignment_params": {"horizontal": "left", "vertical": "center", "wrap_text": True},
        },
    }


def apply_style(cell, style_def, style_type):
    """Apply style to cell using NEW objects (avoids shared object issues)."""
    if "font_params" in style_def:
        cell.font = Font(**style_def["font_params"])
    if "fill_params" in style_def:
        cell.fill = PatternFill(**style_def["fill_params"])
    if "alignment_params" in style_def:
        cell.alignment = Alignment(**style_def["alignment_params"])
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 3: COLUMN DEFINITIONS (BASE + EXTENDED + REGULATORY)
# ============================================================================

def get_base_columns() -> dict:
    """Base columns A-Q (17 columns)."""
    return {
        "Cloud_Service_Name": 30, "Service_Type": 18, "Vendor_Name": 25,
        "Service_Criticality": 18, "Data_Classification": 18, "Data_Residency": 18,
        "Contract_Status": 18, "Status": 15, "Evidence_Location": 30,
        "Gap_Description": 35, "Remediation_Needed": 14, "Exception_ID": 14,
        "Risk_ID": 14, "Compensating_Controls": 30, "Service_Owner": 22,
        "Target_Remediation_Date": 18, "Budget_Impact": 16,
    }


def get_extended_columns() -> dict:
    """Extended columns R-X (7 columns)."""
    return {
        "Monthly_Cost_CHF": 16, "Annual_Contract_Value": 18, "Licensed_Users": 14,
        "Integration_Count": 14, "Backup_Available": 16, "Exit_Feasibility": 16,
        "Last_Review_Date": 16,
    }


def get_regulatory_columns() -> dict:
    """Regulatory columns Y-AG (9 columns) for DORA/NIS2/AI Act/CLOUD Act."""
    return {
        "Provider_HQ_Jurisdiction": 20, "CLOUD_Act_Exposure": 20,
        "Data_Processing_Locations": 25, "EU_Data_Boundary": 16,
        "DORA_In_Scope": 14, "NIS2_In_Scope": 14, "AI_Classification": 18,
        "Concentration_Risk": 18, "Alternatives_Identified": 16,
    }


def get_all_columns() -> dict:
    """All columns A-AG (33 total)."""
    all_cols = {}
    all_cols.update(get_base_columns())
    all_cols.update(get_extended_columns())
    all_cols.update(get_regulatory_columns())
    return all_cols

# ============================================================================
# SECTION 4: VALIDATION DEFINITIONS
# ============================================================================

def create_all_validations(ws) -> dict:
    """Create ALL data validations (base + extended + regulatory)."""
    validations = {}
    
    # --- BASE VALIDATIONS ---
    validations['service_type'] = DataValidation(
        type="list", formula1='"SaaS,IaaS,PaaS,Security,Storage,Collaboration,Other"', allow_blank=True)
    validations['criticality'] = DataValidation(
        type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    validations['data_class'] = DataValidation(
        type="list", formula1='"Restricted,Confidential,Internal,Public,Mixed,N/A"', allow_blank=True)
    validations['residency'] = DataValidation(
        type="list", formula1='"Switzerland,EU,USA,Global,Unknown"', allow_blank=True)
    validations['contract_status'] = DataValidation(
        type="list", formula1='"Active,Renewal Due,Expired,Under Negotiation,Trial"', allow_blank=True)
    validations['status'] = DataValidation(
        type="list", formula1=f'"{CHECK} Compliant,⚠️ Partial,❌ Non-Compliant,N/A"', allow_blank=True)
    validations['yes_no'] = DataValidation(
        type="list", formula1='"Yes,No"', allow_blank=True)
    validations['budget'] = DataValidation(
        type="list", formula1='"High (>$50K),Medium ($10K-$50K),Low (<$10K),None"', allow_blank=True)
    
    # --- EXTENDED VALIDATIONS ---
    validations['backup'] = DataValidation(
        type="list", formula1='"Yes,No,Planned,Unknown"', allow_blank=True)
    validations['exit_feasibility'] = DataValidation(
        type="list", formula1='"Easy (≤30 days),Medium (31-90 days),Hard (>90 days),Unknown"', allow_blank=True)
    
    # --- REGULATORY VALIDATIONS ---
    validations['hq_jurisdiction'] = DataValidation(
        type="list", formula1='"' + ','.join(PROVIDER_HQ_JURISDICTION) + '"', allow_blank=True)
    validations['cloud_act'] = DataValidation(
        type="list", formula1='"' + ','.join(CLOUD_ACT_EXPOSURE) + '"', allow_blank=True)
    validations['eu_boundary'] = DataValidation(
        type="list", formula1='"' + ','.join(EU_DATA_BOUNDARY_OPTIONS) + '"', allow_blank=True)
    validations['dora_scope'] = DataValidation(
        type="list", formula1='"' + ','.join(YES_NO_NOT_DETERMINED) + '"', allow_blank=True)
    validations['nis2_scope'] = DataValidation(
        type="list", formula1='"' + ','.join(YES_NO_NOT_DETERMINED) + '"', allow_blank=True)
    validations['ai_class'] = DataValidation(
        type="list", formula1='"' + ','.join(AI_RISK_CLASSIFICATION) + '"', allow_blank=True)
    validations['concentration'] = DataValidation(
        type="list", formula1='"' + ','.join(CONCENTRATION_RISK_LEVEL) + '"', allow_blank=True)
    validations['alternatives'] = DataValidation(
        type="list", formula1='"' + ','.join(YES_NO_PARTIAL) + '"', allow_blank=True)
    
    return validations


def finalize_validations(ws, validations):
    """Add only data validations that have cells assigned to avoid Excel repair."""
    for dv in validations.values():
        if dv.sqref:
            ws.add_data_validation(dv)


def apply_validations_to_range(ws, validations: dict, start_row: int, end_row: int):
    """Apply validations to data entry range."""
    for row in range(start_row, end_row + 1):
        # Base columns
        validations['service_type'].add(ws.cell(row=row, column=2))    # B
        validations['criticality'].add(ws.cell(row=row, column=4))     # D
        validations['data_class'].add(ws.cell(row=row, column=5))      # E
        validations['residency'].add(ws.cell(row=row, column=6))       # F
        validations['contract_status'].add(ws.cell(row=row, column=7)) # G
        validations['status'].add(ws.cell(row=row, column=8))          # H
        validations['yes_no'].add(ws.cell(row=row, column=11))         # K
        validations['budget'].add(ws.cell(row=row, column=17))         # Q
        
        # Extended columns
        validations['backup'].add(ws.cell(row=row, column=22))         # V
        validations['exit_feasibility'].add(ws.cell(row=row, column=23)) # W
        
        # Regulatory columns (Y=25, Z=26, AA=27, AB=28, AC=29, AD=30, AE=31, AF=32, AG=33)
        validations['hq_jurisdiction'].add(ws.cell(row=row, column=25))  # Y
        validations['cloud_act'].add(ws.cell(row=row, column=26))        # Z
        # AA (27) = free text, no validation
        validations['eu_boundary'].add(ws.cell(row=row, column=28))      # AB
        validations['dora_scope'].add(ws.cell(row=row, column=29))       # AC
        validations['nis2_scope'].add(ws.cell(row=row, column=30))       # AD
        validations['ai_class'].add(ws.cell(row=row, column=31))         # AE
        validations['concentration'].add(ws.cell(row=row, column=32))    # AF
        validations['alternatives'].add(ws.cell(row=row, column=33))     # AG


# ============================================================================
# SECTION 5: CHECKLIST DEFINITIONS
# ============================================================================

def get_checklist_items() -> list:
    """Return combined checklist (base + regulatory)."""
    base = [
        ("All cloud services documented in CMDB", "CMDB export"),
        ("Data classification assigned", "Data classification matrix"),
        ("Service criticality assessed (BIA)", "Business impact analysis"),
        ("Service owner identified", "RACI matrix"),
        ("Contract status tracked", "Contract management system"),
        ("Monthly costs documented", "Finance system"),
        ("User license count verified", "License report"),
        ("Integration dependencies mapped", "Architecture diagram"),
        ("Backup strategy defined", "BC/DR plan"),
        ("Exit feasibility evaluated", "Exit strategy document"),
        ("Data residency verified", "Vendor DPA"),
        ("Last review date recorded", "Review calendar"),
    ]
    regulatory = [
        ("Provider HQ jurisdiction documented", "Vendor documentation"),
        ("CLOUD Act exposure assessed", "Risk assessment"),
        ("Data processing locations verified", "DPA, vendor statement"),
        ("DORA applicability determined", "Regulatory review"),
        ("NIS2 applicability determined", "Regulatory review"),
        ("AI system classification assessed", "AI inventory"),
        ("Concentration risk evaluated", "Risk assessment"),
        ("Alternative providers identified", "Market analysis"),
    ]
    return base + regulatory

# ============================================================================
# SECTION 6: INSTRUCTIONS SHEET
# ============================================================================

def create_instructions_sheet(ws):
    """Create Instructions & Legend sheet with regulatory compliance info."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header — single merged row with two-line title
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.5.23.S1  -  Cloud Service Inventory & Classification\n"
        "ISO/IEC 27001:2022 - Control A.5.23: Information Security for Use of Cloud Services"
    )
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].border = border
    ws.row_dimensions[1].height = 40

    # Document Information
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.5.23.S1"),
        ("Assessment Area", "Cloud Service Inventory & Classification"),
        ("Related Policy", "ISMS-POL-A.5.19-23-S1, S5"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        row += 1

    # Regulatory Compliance Section
    row += 1
    ws[f"A{row}"] = "REGULATORY COMPLIANCE"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="003366")

    row += 1
    reg_info = [
        "This workbook includes regulatory compliance tracking for:",
        f"{BULLET} DORA (Digital Operational Resilience Act) - EU financial entity requirements",
        f"{BULLET} NIS2 (Network and Information Security Directive 2) - Essential/important entities",
        f"{BULLET} EU AI Act - AI system classification and provider evaluation",
        f"{BULLET} US CLOUD Act - Jurisdictional risk for US-headquartered providers",
        "",
        "Regulatory columns (Y-AG) are included in each assessment sheet.",
        "Complete only the regulatory sections applicable to your organisation.",
    ]
    for line in reg_info:
        ws[f"A{row}"] = line
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        row += 1

    # How to Use This Workbook
    row += 1
    ws[f"A{row}"] = "Instructions"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Complete each assessment sheet for all cloud services in your inventory.",
        "2. Use dropdown menus for standardised entries.",
        "3. Fill in yellow-highlighted cells with your service-specific information.",
        "4. Complete regulatory columns (Y-AG) for DORA/NIS2/AI Act/CLOUD Act as applicable.",
        "5. Provide evidence location for each inventory entry.",
        "6. Classify each service by deployment model, data sensitivity, and business criticality.",
        "7. Update the inventory when cloud services are added, changed, or decommissioned.",
        "8. Summary Dashboard auto-calculates compliance statistics.",
        "9. Maintain the Evidence Register with all supporting documentation.",
        "10. Obtain final approval and sign-off from Cloud Security Owner and CISO.",
    ]
    row += 1
    for instr in instructions:
        ws[f"A{row}"] = instr
        row += 1

    # Status Legend — proper table with headers
    row += 1
    ws[f"A{row}"] = "Status Legend"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    for ci, hdr in enumerate(("Symbol", "Status", "Description"), start=1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    legend = [
        (CHECK, "Compliant", "Service meets all security and classification requirements", "C6EFCE"),
        (WARNING, "Partial", "Some requirements met, gaps exist (e.g., missing classification)", "FFEB9C"),
        (XMARK, "Non-Compliant", "Service does not meet minimum requirements", "FFC7CE"),
        ("—", "Not Applicable", "Requirement does not apply to this service", None),
    ]

    row += 1
    for sym, status, desc, color in legend:
        ws.cell(row=row, column=1, value=sym).border = border
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if color:
            s.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

    # Acceptable Evidence
    row += 1
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    evidence_types = [
        "✓ Cloud service contracts and terms of service",
        "✓ Data Processing Agreement (DPA) / Addendum",
        "✓ Service Level Agreement (SLA)",
        "✓ ISO 27001 / ISO 27017 / ISO 27018 certificate (current)",
        "✓ SOC 2 Type II report (within last 12 months)",
        "✓ CSA STAR registration or certification",
        "✓ GDPR/nFADP compliance attestation",
        "✓ Data residency and sovereignty documentation",
        "✓ Cloud service architecture diagrams",
        "✓ Business impact assessment for each cloud service",
        "✓ Risk assessment documentation",
        "✓ Shared responsibility model documentation",
        "✓ Exit strategy and data portability plan",
        "✓ Encryption and key management documentation",
        "✓ Jurisdictional risk assessment (for US-nexus providers)",
    ]
    row += 1
    for ev in evidence_types:
        ws[f"A{row}"] = ev
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 7: ASSESSMENT SHEET CREATOR
# ============================================================================

def create_assessment_sheet(ws, styles, title, policy_req, question, checklist, subtitle=None):
    """Create a service assessment sheet with all 33 columns."""
    columns = get_all_columns()
    total_cols = len(columns)
    last_col = get_column_letter(total_cols)

    # Header
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"{title}\n{policy_req}"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 50

    # Subtitle (row 2)
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = subtitle or f"{title} | Assessment sheet for ISO 27001:2022 A.5.23 compliance"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10)
    ws["A2"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Question
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = question
    ws["A3"].font = Font(bold=True, size=11)
    
    # Response dropdown
    ws["A4"] = "Response:"
    ws["B4"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws["B4"].border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    resp_dv = DataValidation(type="list", formula1='"Yes,No,Partial,Under Review"', allow_blank=False)
    ws.add_data_validation(resp_dv)
    resp_dv.add(ws["B4"])
    
    # Column headers (row 6)
    header_row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        apply_style(cell, styles["column_header"], "header")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # Data entry area (rows 8-58: MAX_DATA_ROWS standard = 51 rows)
    start_row, end_row = 8, 58
    for row in range(start_row, end_row + 1):
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Apply validations
    validations = create_all_validations(ws)
    apply_validations_to_range(ws, validations, start_row, end_row)
    
    # Example row
    ws["A8"] = "Example: Microsoft 365"
    ws["B8"] = "SaaS"
    ws["C8"] = "Microsoft"
    ws["D8"] = "Critical"
    ws["H8"] = f"{CHECK} Compliant"
    ws["Y8"] = "United States"
    ws["Z8"] = "Mitigated (EU Data Boundary)"
    ws["AC8"] = "Yes"
    ws["AD8"] = "Yes"
    ws["AF8"] = "High (Few alternatives)"
    ws["A8"].font = Font(italic=True, color="666666")
    
    # Checklist (starts 3 rows after data ends)
    checklist_row = end_row + 3
    ws.merge_cells(f"A{checklist_row}:D{checklist_row}")
    ws[f"A{checklist_row}"] = "COMPLIANCE CHECKLIST"
    apply_style(ws[f"A{checklist_row}"], styles["subheader"], "subheader")
    
    checklist_row += 2
    headers = ["Item", "Requirement", "Status", "Evidence"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=checklist_row, column=col_idx, value=h)
        apply_style(cell, styles["column_header"], "header")
    
    status_dv = DataValidation(type="list", formula1=f'"{CHECK},{WARNING},{XMARK}"', allow_blank=True)
    ws.add_data_validation(status_dv)
    
    checklist_row += 1
    for req, evidence in checklist:
        ws.cell(row=checklist_row, column=1, value="☐")
        ws.cell(row=checklist_row, column=2, value=req)
        status_dv.add(ws.cell(row=checklist_row, column=3))
        ws.cell(row=checklist_row, column=3).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=checklist_row, column=3).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws.cell(row=checklist_row, column=4, value=evidence)
        checklist_row += 1
    
    ws.freeze_panes = "A7"
    finalize_validations(ws, validations)


# ============================================================================
# SECTION 7B: EXIT STRATEGY ASSESSMENT SHEET (SHEET 4)
# ============================================================================

def create_exit_strategy_sheet(ws, styles):
    """
    Create Sheet 4: Exit Feasibility & Strategy Assessment
    
    Implements POL-A.5.19-23-S5 Section 8.1.1 (Exit Strategy Framework)
    
    Column Structure:
    - Columns A-Q: Basic service information (17 columns)
    - Columns R-AC: Exit Strategy framework (12 columns)
    Total: 29 columns
    """
    
    # === HEADER ===
    ws.merge_cells('A1:AC1')
    ws['A1'] = "SHEET 4: EXIT FEASIBILITY & STRATEGY ASSESSMENT"
    ws['A1'].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    
    # === POLICY REFERENCE ===
    ws.merge_cells('A2:AC2')
    ws['A2'] = "POL-S5 Section 8.1.1: Cloud-to-Cloud (90%+), Hybrid (5-10%), On-Premises (<5% only when justified)"
    ws['A2'].font = Font(name="Calibri", size=9, italic=True)
    ws['A2'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # === COLUMN HEADERS (Row 3) ===
    headers = [
        # Columns A-Q: Basic Service Information
        ("A", "Service Name", 25),
        ("B", "Provider", 20),
        ("C", "Service Type", 18),
        ("D", "Environment", 15),
        ("E", "Business Owner", 20),
        ("F", "Technical Owner", 20),
        ("G", "Data Classification", 18),
        ("H", "Criticality", 15),
        ("I", "Users", 12),
        ("J", "Cost (Annual CHF)", 18),
        ("K", "Contract End Date", 15),
        ("L", "Primary Region", 18),
        ("M", "Backup Region", 18),
        ("N", "Integration Count", 15),
        ("O", "API Dependency", 15),
        ("P", "Compliance Certs", 20),
        ("Q", "Current Status", 15),
        
        # Columns R-AC: Exit Strategy Framework
        ("R", "Exit Strategy Type", 20),
        ("S", "Alternative Identified", 22),
        ("T", "Export Format Available", 22),
        ("U", "Export Tested", 16),
        ("V", "Data Volume (GB)", 16),
        ("W", "Migration Complexity", 24),
        ("X", "Lock-In Risk", 16),
        ("Y", "Hybrid: Workload Segmentation", 28),
        ("Z", "Hybrid: Data Sync Latency", 28),
        ("AA", "OnPrem: TCO Analysis Complete", 28),
        ("AB", "OnPrem: Infrastructure Available", 30),
        ("AC", "OnPrem: Staffing Plan Documented", 30),
    ]
    
    thin = Side(style="thin")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
        ws.column_dimensions[col].width = width
    
    # === DATA VALIDATION SETUP ===
    
    # Basic validations
    val_service_type = DataValidation(
        type="list", formula1='"SaaS,IaaS,PaaS,Security,Storage,Other"', allow_blank=False)
    val_environment = DataValidation(
        type="list", formula1='"Production,Development,Test,Staging,All"', allow_blank=False)
    val_data_class = DataValidation(
        type="list", formula1='"Restricted,Confidential,Internal,Public,Mixed"', allow_blank=False)
    val_criticality = DataValidation(
        type="list", formula1='"Critical,High,Medium,Low"', allow_blank=False)
    val_api_dependency = DataValidation(
        type="list", formula1='"Yes,No,Partial"', allow_blank=False)
    val_status = DataValidation(
        type="list", formula1=f'"{CHECK} Compliant,⚠️ Partial,❌Non-Compliant"', allow_blank=False)
    
    # Exit Strategy validations
    val_exit_strategy = DataValidation(
        type="list", formula1='"' + ','.join(EXIT_STRATEGY_TYPE) + '"', allow_blank=False)
    val_exit_strategy.error = "Select: Cloud-to-Cloud (default), Hybrid, On-Premises, or Not Yet Determined"
    val_exit_strategy.errorTitle = "Invalid Exit Strategy"
    
    val_alternative = DataValidation(
        type="list", formula1='"' + ','.join(ALTERNATIVE_IDENTIFIED) + '"', allow_blank=False)
    
    val_export_format = DataValidation(
        type="list", formula1='"' + ','.join(EXPORT_FORMAT) + '"', allow_blank=False)
    
    val_export_tested = DataValidation(
        type="list", formula1='"' + ','.join(EXPORT_TESTED) + '"', allow_blank=False)
    
    val_data_volume = DataValidation(
        type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    val_data_volume.error = "Enter a positive number (GB)"
    val_data_volume.errorTitle = "Invalid Data Volume"
    
    val_migration_complexity = DataValidation(
        type="list", formula1='"' + ','.join(MIGRATION_COMPLEXITY) + '"', allow_blank=False)
    
    val_lock_in_risk = DataValidation(
        type="list", formula1='"' + ','.join(LOCK_IN_RISK) + '"', allow_blank=False)
    
    val_hybrid_workload = DataValidation(
        type="list", formula1='"' + ','.join(HYBRID_WORKLOAD_SEGMENTATION) + '"', allow_blank=False)
    
    val_hybrid_sync = DataValidation(
        type="list", formula1='"' + ','.join(HYBRID_DATA_SYNC_LATENCY) + '"', allow_blank=False)
    
    val_onprem_tco = DataValidation(
        type="list", formula1='"' + ','.join(ONPREM_TCO_ANALYSIS) + '"', allow_blank=False)
    
    val_onprem_infra = DataValidation(
        type="list", formula1='"' + ','.join(ONPREM_INFRASTRUCTURE) + '"', allow_blank=False)
    
    val_onprem_staffing = DataValidation(
        type="list", formula1='"' + ','.join(ONPREM_STAFFING_PLAN) + '"', allow_blank=False)
    
    # Add validations to worksheet
    for val in [val_service_type, val_environment, val_data_class, val_criticality, val_api_dependency, val_status,
                val_exit_strategy, val_alternative, val_export_format, val_export_tested, val_data_volume,
                val_migration_complexity, val_lock_in_risk, val_hybrid_workload, val_hybrid_sync,
                val_onprem_tco, val_onprem_infra, val_onprem_staffing]:
        ws.add_data_validation(val)
    
    # === SAMPLE ROW (row 4) — F2F2F2 grey with realistic example data ===
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    _all_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q',
                 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']
    _sample_exit = {
        'A': "Microsoft 365", 'B': "Microsoft", 'C': "SaaS", 'D': "Production",
        'E': "IT Director", 'F': "Cloud Admin", 'G': "Confidential", 'H': "Critical",
        'I': "250", 'J': "125000", 'K': "31.12.2026", 'L': "West Europe (Amsterdam)",
        'M': "North Europe (Dublin)", 'N': "15", 'O': "Yes",
        'P': "ISO 27001, SOC 2 Type II", 'Q': f"{CHECK} Compliant",
        'R': "Cloud-to-Cloud", 'S': "Cloud Provider (GCP)",
        'T': "Standard (CSV/JSON)", 'U': "Yes", 'V': "2500",
        'W': "Cloud-to-Cloud (Medium)", 'X': "Medium",
        'Y': "Not Applicable", 'Z': "Not Applicable",
        'AA': "N/A", 'AB': "N/A", 'AC': "N/A",
    }
    for col in _all_cols:
        cell = ws[f'{col}4']
        cell.value = _sample_exit.get(col, "")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(italic=True, color="666666")
        cell.border = border
    val_service_type.add(ws['C4'])
    val_environment.add(ws['D4'])
    val_data_class.add(ws['G4'])
    val_criticality.add(ws['H4'])
    val_api_dependency.add(ws['O4'])
    val_status.add(ws['Q4'])
    val_exit_strategy.add(ws['R4'])
    val_alternative.add(ws['S4'])
    val_export_format.add(ws['T4'])
    val_export_tested.add(ws['U4'])
    val_data_volume.add(ws['V4'])
    val_migration_complexity.add(ws['W4'])
    val_lock_in_risk.add(ws['X4'])
    val_hybrid_workload.add(ws['Y4'])
    val_hybrid_sync.add(ws['Z4'])
    val_onprem_tco.add(ws['AA4'])
    val_onprem_infra.add(ws['AB4'])
    val_onprem_staffing.add(ws['AC4'])

    # === APPLY VALIDATIONS TO DATA ROWS (5-54: 50 empty FFFFCC rows) ===
    for row in range(5, 55):
        # Basic columns
        val_service_type.add(ws[f'C{row}'])
        val_environment.add(ws[f'D{row}'])
        val_data_class.add(ws[f'G{row}'])
        val_criticality.add(ws[f'H{row}'])
        val_api_dependency.add(ws[f'O{row}'])
        val_status.add(ws[f'Q{row}'])

        # Exit Strategy columns
        val_exit_strategy.add(ws[f'R{row}'])
        val_alternative.add(ws[f'S{row}'])
        val_export_format.add(ws[f'T{row}'])
        val_export_tested.add(ws[f'U{row}'])
        val_data_volume.add(ws[f'V{row}'])
        val_migration_complexity.add(ws[f'W{row}'])
        val_lock_in_risk.add(ws[f'X{row}'])
        val_hybrid_workload.add(ws[f'Y{row}'])
        val_hybrid_sync.add(ws[f'Z{row}'])
        val_onprem_tco.add(ws[f'AA{row}'])
        val_onprem_infra.add(ws[f'AB{row}'])
        val_onprem_staffing.add(ws[f'AC{row}'])

        # Yellow fill + borders for input cells
        for col in _all_cols:
            cell = ws[f'{col}{row}']
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
    
    # === CONDITIONAL FORMATTING - WARNINGS ===
    from openpyxl.formatting.rule import FormulaRule

    # WARNING 1: On-Premises strategy → Orange fill (warn: <5% justified)
    onprem_warning = FormulaRule(
        formula=['$R4="On-Premises"'],
        fill=PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        stopIfTrue=False
    )
    ws.conditional_formatting.add('R4:AC54', onprem_warning)
    
    # WARNING 2: Critical Lock-In Risk → Red fill
    critical_lock_in = FormulaRule(
        formula=['$X4="Critical"'],
        fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        font=Font(color="FFFFFF", bold=True),
        stopIfTrue=False
    )
    ws.conditional_formatting.add('X4:X54', critical_lock_in)
    
    # WARNING 3: Export not tested for Critical services → Yellow fill
    untested_critical = FormulaRule(
        formula=['AND($U4="No", $H4="Critical")'],
        fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        stopIfTrue=False
    )
    ws.conditional_formatting.add('U4:U54', untested_critical)
    
    # === CHECKLIST SECTION (Row 57+) ===
    ws.merge_cells('A57:AC57')
    ws['A57'] = "EXIT STRATEGY COMPLIANCE CHECKLIST (22 items)"
    ws['A57'].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws['A57'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws['A57'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[57].height = 30
    
    checklist_items = [
        # Original 15 items
        f"{CHECK} Exit strategy documented for Critical/High services",
        f"{CHECK} Exit triggers defined",
        f"{CHECK} Alternative providers/solutions identified",
        f"{CHECK} Alternative provider assessment completed",
        f"{CHECK} Migration effort estimated",
        f"{CHECK} Data extraction requirements documented",
        f"{CHECK} Application dependencies mapped",
        f"{CHECK} Integration points documented",
        f"{CHECK} Lock-in risks identified and documented",
        f"{CHECK} Exit cost estimate completed",
        f"{CHECK} Exit timeline estimate completed",
        f"{CHECK} Data export format verified",
        f"{CHECK} Migration testing requirements defined",
        f"{CHECK} Exit plan reviewed and approved",
        f"{CHECK} Exit strategy included in vendor contracts",
        
        # 7 items for Exit Strategy framework
        f"{CHECK} Cloud-to-Cloud alternatives evaluated (min. 2 providers)",
        f"{CHECK} Hybrid workload segmentation documented (if applicable)",
        f"{CHECK} On-Premises TCO analysis completed (if applicable)",
        f"{CHECK} Annual PoC testing completed (DORA Art. 28.6 for Critical)",
        f"{CHECK} Vendor lock-in risks assessed and mitigation documented",
        f"{CHECK} Migration cost estimates align with exit strategy type",
        f"{CHECK} Exit strategy type justified with evidence (esp. On-Premises)",
    ]
    
    row = 106
    for idx, item in enumerate(checklist_items, 1):
        ws[f'A{row}'] = f"{idx}. {item}"
        ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    # === HELP TEXT SECTION (Row 135+) ===
    ws.merge_cells('A135:AC135')
    ws['A135'] = "EXIT STRATEGY GUIDANCE - POLICY REQUIREMENTS"
    ws['A135'].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws['A135'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws['A135'].alignment = Alignment(horizontal="center", vertical="center")

    help_text = [
        "",
        f"{CHECK} CLOUD-TO-CLOUD (90%+ of services - DEFAULT STRATEGY)",
        "   • Migrate to alternative cloud provider (AWS, Azure, GCP, OVHcloud, Alibaba, etc.)",
        "   • Timeline: 3-6 months typical",
        "   • Cost: CHF 40K-200K one-time",
        "   • Benefits: Zero CAPEX, maintained elasticity, modern cloud-native capabilities",
        "   • Required: Min. 2 alternative providers identified, annual PoC testing (DORA Art. 28.6)",
        "",
        f"{WARNING} HYBRID (5-10% of services - PARTIAL REPATRIATION)",
        "   • Some workloads in cloud, others on-premises (e.g., app in cloud, DB on-prem)",
        "   • Timeline: 6-12 months typical",
        "   • Cost: CHF 335K-1.6M (3-year TCO)",
        "   • Use cases: Data sovereignty (FADP), cost optimization, latency requirements",
        "   • Required: Workload segmentation documented, data sync latency assessed, network design",
        "",
        f"{WARNING}{WARNING} ON-PREMISES (<5% of services - FULL BUILD-BACK)",
        "   • Complete migration to [Organisation]-owned infrastructure",
        "   • Timeline: 6-18 months (long!)",
        "   • Cost: CHF 2.94M-7.04M (5-year TCO)",
        "   • Justification: Regulatory mandate, cost inversion (>CHF 500K/yr cloud), strategic independence",
        "   • Required: TCO analysis, infrastructure capacity plan, staffing plan (3-10 FTEs)",
        "   • WARNING: Economically justified in <5% of cases - validate TCO carefully!",
        "",
        f"{LOCK} VENDOR LOCK-IN RISK LEVELS:",
        "   • Low: Standard protocols, open formats, portable architecture",
        "   • Medium: Some proprietary features, manageable migration effort",
        "   • High: Deep integration with proprietary services, significant refactoring needed",
        "   • Critical: Mission-critical dependency on unique capability, no viable alternatives",
        "",
        f"{SHIELD} DORA COMPLIANCE (Financial Sector):",
        "   • Article 28.6: Annual PoC testing of exit strategies for critical ICT providers",
        "   • Article 28.9: Concentration risk mitigation (reduce dependency on single hyperscaler)",
        "   • Track PoC testing in ISMS-IMP-A.5.23.S4 (Governance workbook)",
        "",
        "For detailed cost models and TCO calculations, refer to:",
        "   • POL-S5 Section 8.1.1.1 (Cloud-to-Cloud)",
        "   • POL-S5 Section 8.1.1.2 (Hybrid)",
        "   • POL-S5 Section 8.1.1.3 (On-Premises)",
        "",
    ]
    
    row = 136
    for line in help_text:
        ws[f'A{row}'] = line
        ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    ws.freeze_panes = "A4"

    logger.info("   [OK] Sheet 4: Exit Feasibility & Strategy Assessment created")
    logger.info(f"        Columns: A-AC (29 total)")
    logger.info(f"        Exit Strategy framework: Columns R-AC (12 columns)")
    logger.info(f"        Conditional formatting: On-Prem warning, Lock-In risk, Untested exports")


# ============================================================================
# SECTION 6: DATA CLASSIFICATION SHEET
# ============================================================================

def create_data_classification_sheet(ws, styles):
    """Create Data Classification sheet for cloud-stored data inventory."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "DATA CLASSIFICATION - CLOUD STORAGE INVENTORY"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 30

    # Instruction row
    ws.merge_cells("A2:J2")
    ws["A2"] = "Classify all data types stored in cloud services. Required for ISO 27001 A.5.12 (Classification), A.5.13 (Labelling), and GDPR Article 30."
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)

    # Headers
    headers = [
        ("A", "Data_ID", 12),
        ("B", "Data Category", 25),
        ("C", "Description", 35),
        ("D", "Classification Level", 18),
        ("E", "Primary Cloud Service", 22),
        ("F", "Storage Location", 22),
        ("G", "Data Owner", 18),
        ("H", "Retention Period", 16),
        ("I", "Personal Data (GDPR)", 18),
        ("J", "Cross-Border Transfer", 18),
    ]

    thin = Side(style="thin")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
        ws.column_dimensions[col].width = width

    # Data validations
    val_classification = DataValidation(
        type="list", formula1='"Restricted,Confidential,Internal,Public"', allow_blank=False)
    val_personal_data = DataValidation(
        type="list", formula1='"Yes - Special Category,Yes - Personal,No"', allow_blank=False)
    val_cross_border = DataValidation(
        type="list", formula1='"None,EEA Only,Adequacy Decision,SCCs,Not Assessed"', allow_blank=False)

    ws.add_data_validation(val_classification)
    ws.add_data_validation(val_personal_data)
    ws.add_data_validation(val_cross_border)

    # Sample row (row 4) — F2F2F2 grey with realistic example data
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sample_data = [
        "DC-001",
        "Customer PII",
        "Customer names, email addresses, phone numbers in CRM system",
        "Confidential",
        "AWS S3",
        "eu-west-1 (Ireland)",
        "VP Sales",
        "7 years post-contract",
        "Yes - Personal",
        "EEA Only"
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(italic=True, color="666666")
        cell.border = border

    val_classification.add(ws['D4'])
    val_personal_data.add(ws['I4'])
    val_cross_border.add(ws['J4'])

    # Empty rows 5-54 (50 empty rows = 51 total with sample)
    for row in range(5, 55):
        for col_idx in range(1, 11):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

        val_classification.add(ws[f'D{row}'])
        val_personal_data.add(ws[f'I{row}'])
        val_cross_border.add(ws[f'J{row}'])

    ws.freeze_panes = "A4"
    logger.info("   [OK] Sheet 6: Data Classification created (51 rows: 1 sample + 50 empty)")


# ============================================================================
# SECTION 7: SERVICE CRITICALITY SHEET
# ============================================================================

def create_service_criticality_sheet(ws, styles):
    """Create Service Criticality Assessment sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "SERVICE CRITICALITY ASSESSMENT"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 30

    # Instruction row
    ws.merge_cells("A2:L2")
    ws["A2"] = "Assess criticality of each cloud service for BIA and exit strategy planning. Required for ISO 27001 A.5.29 (Business Continuity) and DORA Article 28."
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)

    # Headers
    headers = [
        ("A", "Service_ID", 12),
        ("B", "Cloud Service Name", 28),
        ("C", "Service Category", 18),
        ("D", "Business Process Supported", 28),
        ("E", "Criticality Level", 16),
        ("F", "RTO (Hours)", 14),
        ("G", "RPO (Hours)", 14),
        ("H", "MTPD (Hours)", 14),
        ("I", "Single Point of Failure", 18),
        ("J", "Workaround Available", 18),
        ("K", "DORA Scope", 14),
        ("L", "Exit Priority", 14),
    ]

    thin = Side(style="thin")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
        ws.column_dimensions[col].width = width

    # Data validations
    val_category = DataValidation(
        type="list", formula1='"SaaS,IaaS,PaaS,Security,Storage,Other"', allow_blank=False)
    val_criticality = DataValidation(
        type="list", formula1='"Critical,High,Medium,Low"', allow_blank=False)
    val_yesno = DataValidation(
        type="list", formula1='"Yes,No,Partial"', allow_blank=False)
    val_dora = DataValidation(
        type="list", formula1='"In-Scope,Out-of-Scope,TBD"', allow_blank=False)
    val_priority = DataValidation(
        type="list", formula1='"P1 - Critical,P2 - High,P3 - Medium,P4 - Low"', allow_blank=False)

    ws.add_data_validation(val_category)
    ws.add_data_validation(val_criticality)
    ws.add_data_validation(val_yesno)
    ws.add_data_validation(val_dora)
    ws.add_data_validation(val_priority)

    # Sample row (row 4) — F2F2F2 grey with realistic example data
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sample_data = [
        "SC-001",
        "Microsoft 365 (Email, SharePoint, Teams)",
        "SaaS",
        "Email communication, document collaboration, internal comms",
        "Critical",
        "4",
        "1",
        "8",
        "Yes",
        "Partial",
        "In-Scope",
        "P1 - Critical"
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(italic=True, color="666666")
        cell.border = border

    val_category.add(ws['C4'])
    val_criticality.add(ws['E4'])
    val_yesno.add(ws['I4'])
    val_yesno.add(ws['J4'])
    val_dora.add(ws['K4'])
    val_priority.add(ws['L4'])

    # Empty rows 5-54 (50 empty rows = 51 total with sample)
    for row in range(5, 55):
        for col_idx in range(1, 13):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

        val_category.add(ws[f'C{row}'])
        val_criticality.add(ws[f'E{row}'])
        val_yesno.add(ws[f'I{row}'])
        val_yesno.add(ws[f'J{row}'])
        val_dora.add(ws[f'K{row}'])
        val_priority.add(ws[f'L{row}'])

    ws.freeze_panes = "A4"
    logger.info("   [OK] Sheet 7: Service Criticality Assessment created (51 rows: 1 sample + 50 empty)")


# ============================================================================
# SECTION 8: SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with compliance and regulatory metrics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "CLOUD SERVICE INVENTORY - COMPLIANCE DASHBOARD"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 35
    
    # Compliance by category
    row = 3
    ws[f"A{row}"] = "TABLE 1: COMPLIANCE BY SERVICE CATEGORY"
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    
    row += 1
    headers = ["Category", "Total", f"{CHECK} Compliant", f"{WARNING} Partial", f"{XMARK} Non-Compliant", "N/A", "Compliance %"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    categories = [
        ("SaaS Services", "2. SaaS Services"),
        ("IaaS/PaaS", "3. IaaS PaaS Services"),
        ("Storage Services", "5. Cloud Storage Services"),
    ]
    
    row += 1
    start_data_row = row
    data_last_row = 100  # FML-004: Extended from 50 to cover all FFFFCC data rows
    for label, sheet in categories:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=f"=COUNTA('{sheet}'!A8:A{data_last_row})")
        ws.cell(row=row, column=3, value=f'=COUNTIF(\'{sheet}\'!H8:H{data_last_row},"{CHECK}*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF(\'{sheet}\'!H8:H{data_last_row},"{WARNING}*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF(\'{sheet}\'!H8:H{data_last_row},"{XMARK}*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF(\'{sheet}\'!H8:H{data_last_row},"N/A")')
        ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    end_data_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL")
    ws[f"A{row}"].font = Font(bold=True)
    for col in range(2, 7):
        ws.cell(row=row, column=col, value=f'=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{end_data_row})')
        ws.cell(row=row, column=col).font = Font(bold=True)
    ws.cell(row=row, column=7, value=f'=IF(B{row}-F{row}>0,ROUND(C{row}/(B{row}-F{row})*100,1)&"%","N/A")')
    ws.cell(row=row, column=7).font = Font(bold=True)
    row += 1

    # REGULATORY METRICS
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "TABLE 2: REGULATORY COMPLIANCE SUMMARY (DORA / NIS2 / AI Act / CLOUD Act)"
    ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    
    row += 2
    for col_idx, h in enumerate(["Metric", "Count", "Status"], start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    row += 1
    reg_metrics = [
        ("CLOUD Act Potential Exposure", 
         '=COUNTIF(\'2. SaaS Services\'!Z:Z,"Potential*")+COUNTIF(\'3. IaaS PaaS Services\'!Z:Z,"Potential*")'),
        ("DORA In-Scope Services",
         '=COUNTIF(\'2. SaaS Services\'!AC:AC,"Yes")+COUNTIF(\'3. IaaS PaaS Services\'!AC:AC,"Yes")'),
        ("NIS2 In-Scope Services",
         '=COUNTIF(\'2. SaaS Services\'!AD:AD,"Yes")+COUNTIF(\'3. IaaS PaaS Services\'!AD:AD,"Yes")'),
        ("High-Risk AI Systems",
         '=COUNTIF(\'2. SaaS Services\'!AE:AE,"High Risk")+COUNTIF(\'3. IaaS PaaS Services\'!AE:AE,"High Risk")'),
        ("Critical Concentration Risk",
         '=COUNTIF(\'2. SaaS Services\'!AF:AF,"Critical*")+COUNTIF(\'3. IaaS PaaS Services\'!AF:AF,"Critical*")'),
    ]
    
    for metric, formula in reg_metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=3, value=f'=IF(B{row}>0,"{WARNING} Review","{CHECK} OK")')
        row += 1
    
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A3"

# ============================================================================
# SECTION 9: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws):
    """Create Evidence Register sheet."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        ("Evidence ID", 15), ("Assessment Area", 25), ("Evidence Type", 22),
        ("Description", 40), ("Location/Path", 45), ("Date Collected", 16),
        ("Collected By", 20), ("Verification Status", 22),
    ]

    row = 4
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Evidence type dropdown
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Contract,Invoice,CMDB Export,Screenshot,Config File,Risk Assessment,Audit Report,Other"',
        allow_blank=True
    )
    ws.add_data_validation(ev_type_dv)

    # Verification status dropdown
    ver_dv = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=True
    )
    ws.add_data_validation(ver_dv)

    # Sample row with complete example data
    sample_data = [
        "EV-001",
        "Cloud Service Inventory",
        "CMDB Export",
        "Complete inventory of all SaaS/IaaS/PaaS services with vendor details",
        "/evidence/cloud-inventory-q1-2026.xlsx",
        "15.01.2026",
        "IT Security Team",
        "Verified"
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.font = Font(name="Calibri", italic=True, color="666666")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ev_type_dv.add(ws.cell(row=5, column=3))
    ver_dv.add(ws.cell(row=5, column=8))

    # Empty rows 6-104 (99 empty rows = 100 total with sample)
    for r in range(6, 105):
        for col in range(1, 9):
            cell = ws.cell(row=r, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ev_type_dv.add(ws.cell(row=r, column=3))
        ver_dv.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 10: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(ws):
    """Create Approval Sign-Off sheet."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "APPROVAL SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle (row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = "Cloud Service Inventory & Classification | Approval and sign-off for ISO 27001:2022 Control A.5.23"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10)
    ws["A2"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Assessment Summary
    row = 3
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    
    summary = [
        ("Document:", "ISMS-IMP-A.5.23.S1 - Cloud Service Inventory"),
        ("Assessment Period:", ""),
        ("Total Services:", "='8. Summary Dashboard'!B8"),
        ("Overall Compliance:", "='8. Summary Dashboard'!G8"),
        ("Assessment Status:", ""),
    ]
    
    row += 1
    for label, value in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin_s = Side(style="thin")
            ws[f"B{row}"].border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
        row += 1

    # Approval sections
    approvers = [
        ("COMPLETED BY (IT OPERATIONS)", "4472C4"),
        ("REVIEWED BY (ISO)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    
    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
            row += 1
        row += 1
    
    # Final decision
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    thin_fd = Side(style="thin")
    ws[f"B{row}"].border = Border(left=thin_fd, right=thin_fd, top=thin_fd, bottom=thin_fd)

    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin = Side(style="thin")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 11: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution - generates the complete workbook."""
    logger.info("=" * 70)
    logger.info("ISMS-IMP-A.5.23.S1 - Cloud Service Inventory Generator")
    logger.info("WITH REGULATORY COLUMNS (DORA / NIS2 / AI Act / CLOUD Act)")
    logger.info("=" * 70)
    
    wb = create_workbook()
    styles = setup_styles()
    checklist = get_checklist_items()
    
    logger.info("\n[1/10] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"])
    
    logger.info("[2/10] Creating 2. SaaS Services...")
    create_assessment_sheet(
        wb["2. SaaS Services"], styles,
        "SAAS SERVICES INVENTORY",
        "All SaaS applications must be documented with data classification (Policy S5 Section 2.1)",
        "Does your organisation use SaaS applications?",
        checklist,
        subtitle="SaaS Services | Inventory and regulatory assessment of all Software-as-a-Service applications (DORA / NIS2 / AI Act / CLOUD Act)"
    )

    logger.info("[3/10] Creating 3. IaaS PaaS Services...")
    create_assessment_sheet(
        wb["3. IaaS PaaS Services"], styles,
        "IAAS/PAAS SERVICES INVENTORY",
        "All infrastructure services must maintain security baselines (Policy S5 Section 2.2)",
        "Does your organisation use IaaS or PaaS services?",
        checklist,
        subtitle="IaaS/PaaS Services | Inventory and regulatory assessment of infrastructure and platform cloud services"
    )
    
    logger.info("[4/10] Creating 4. Exit Feasibility & Strategy Assessment...")
    create_exit_strategy_sheet(wb["4. Cloud Security Services"], styles)
    
    logger.info("[5/10] Creating 5. Cloud Storage Services...")
    create_assessment_sheet(
        wb["5. Cloud Storage Services"], styles,
        "CLOUD STORAGE SERVICES INVENTORY",
        "All cloud storage must enforce encryption (Policy S5 Section 2.4)",
        "Does your organisation use cloud storage services?",
        checklist,
        subtitle="Cloud Storage Services | Inventory and regulatory assessment of cloud storage and object store solutions"
    )
    
    logger.info("[6/10] Creating 6. Data Classification...")
    create_data_classification_sheet(wb["6. Data Classification"], styles)

    logger.info("[7/10] Creating 7. Service Criticality...")
    create_service_criticality_sheet(wb["7. Service Criticality"], styles)
    
    logger.info("[8/10] Creating 8. Summary Dashboard...")
    create_summary_dashboard(wb["8. Summary Dashboard"], styles)
    
    logger.info("[9/10] Creating 9. Evidence Register...")
    create_evidence_register(wb["9. Evidence Register"])

    logger.info("[10/10] Creating 10. Approval Sign-Off...")
    create_approval_signoff(wb["10. Approval Sign-Of"])
    
    # Save
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.5.23.S1_Inventory_{timestamp}.xlsx"
    _wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
    _wkbk_dir.mkdir(exist_ok=True)
    wb.save(_wkbk_dir / filename)

    logger.info("\n" + "=" * 70)
    logger.info(f"{CHECK} SUCCESS: {filename}")
    logger.info(f"Sheets: {len(wb.sheetnames)}")
    logger.info(f"Exit Strategy: Sheet 4 (A-AC, 29 cols) | Regulatory: DORA/NIS2/AI/CLOUD")
    logger.info("=" * 70)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================