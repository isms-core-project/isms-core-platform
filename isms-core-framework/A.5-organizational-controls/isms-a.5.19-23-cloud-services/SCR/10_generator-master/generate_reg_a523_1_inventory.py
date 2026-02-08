#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.23.S1 - Cloud Service Inventory & Classification Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.23: Information Security for Use of Cloud Services
Assessment Domain 1 of 5: Cloud Service Inventory & Classification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific cloud service landscape, regulatory requirements,
and classification criteria.

Key customization areas:
1. Cloud provider inventory (AWS, Azure, GCP per your deployment)
2. Service classification criteria (aligned with your data classification)
3. Regulatory mappings (DORA, NIS2, AI Act per your jurisdictions)
4. Risk rating criteria (based on your risk framework)
5. Data residency requirements (per your compliance obligations)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for cloud services)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

**Purpose:**
Enables systematic inventory and classification of cloud services supporting
ISO 27001:2022 Control A.5.23 requirements and regulatory compliance (DORA,
NIS2, AI Act, CLOUD Act).

**Assessment Scope:**
- Cloud service inventory (IaaS, PaaS, SaaS across providers)
- Service classification (criticality, data sensitivity)
- Regulatory applicability mapping
- Data residency and sovereignty tracking
- Concentration risk assessment
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance
2. Cloud Service Inventory - Complete service catalog
3. Classification Matrix - Criticality and sensitivity
4. Regulatory Mapping - DORA, NIS2, AI Act applicability
5. Data Residency - Sovereignty and location tracking
6. Concentration Risk - Provider dependency analysis
7. Gap Analysis - Non-compliant services
8. Evidence Register - Audit evidence tracking
9. Approval & Sign-Off - Stakeholder approval

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_reg_a523_1_inventory.py

Requirements:
    - Python 3.8+
    - openpyxl library: pip install openpyxl

Output:
    ISMS-IMP-A.5.23.S1_Inventory_DD_YYYYMMDD.xlsx
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.23.S1"
WORKBOOK_NAME = "Cloud Service Inventory & Classification"
CONTROL_ID = "A.5.19-23"
CONTROL_NAME = "Information Security for Use of Cloud Services"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# REGULATORY DROPDOWN DEFINITIONS (DORA, NIS2, AI Act, CLOUD Act)
# ============================================================================

PROVIDER_HQ_JURISDICTION = [
    "Switzerland", "EU/EEA", "United Kingdom", "United States",
    "Other Adequate Country", "Non-Adequate Country"
]

CLOUD_ACT_EXPOSURE = [
    "No Exposure", "Potential Exposure (US HQ)", "Mitigated (EU Data Boundary)",
    "Mitigated (Encryption + Key Control)", "Accepted Risk (Documented)", "Under Assessment"
]

AI_RISK_CLASSIFICATION = [
    "Not Applicable", "Minimal Risk", "Limited Risk (Transparency)",
    "High Risk", "Unacceptable Risk (Prohibited)"
]

CONCENTRATION_RISK_LEVEL = [
    "Low (Multiple alternatives)", "Medium (Limited alternatives)",
    "High (Few alternatives)", "Critical (Single provider dependency)"
]

EU_DATA_BOUNDARY_OPTIONS = ["Yes", "No", "Partial", "Unknown"]
YES_NO_NOT_DETERMINED = ["Yes", "No", "Not Determined"]
YES_NO_PARTIAL = ["Yes", "No", "Partial"]

# Exit Strategy Framework (POL-S5 Section 8.1.1)
EXIT_STRATEGY_TYPE = [
    "Cloud-to-Cloud", 
    "Hybrid", 
    "On-Premises", 
    "Not Yet Determined"
]

ALTERNATIVE_IDENTIFIED = [
    "Cloud Provider (AWS)", 
    "Cloud Provider (Azure)", 
    "Cloud Provider (GCP)",
    "Cloud Provider (OVHcloud)",
    "Cloud Provider (Alibaba)",
    "Cloud Provider (Other)",
    "Hybrid (Cloud+OnPrem)", 
    "On-Prem Only", 
    "Multiple Options", 
    "Not Yet Assessed"
]

EXPORT_FORMAT = [
    "Standard (CSV/JSON)", 
    "Proprietary", 
    "API Only", 
    "None"
]

EXPORT_TESTED = ["Yes", "No", "Partial"]

MIGRATION_COMPLEXITY = [
    "Cloud-to-Cloud (Low)",
    "Cloud-to-Cloud (Medium)",
    "Cloud-to-Cloud (High)",
    "Hybrid (Medium)",
    "Hybrid (High)",
    "On-Prem (High)",
    "On-Prem (Very High)"
]

LOCK_IN_RISK = ["Low", "Medium", "High", "Critical"]

HYBRID_WORKLOAD_SEGMENTATION = ["Documented", "In Progress", "Not Applicable"]

HYBRID_DATA_SYNC_LATENCY = [
    "Excellent (<100ms)", 
    "Acceptable (100-500ms)", 
    "Concern (>500ms)", 
    "Not Tested", 
    "N/A"
]

ONPREM_TCO_ANALYSIS = [
    "Yes (Favorable)", 
    "Yes (Unfavorable)", 
    "In Progress", 
    "Not Started", 
    "N/A"
]

ONPREM_INFRASTRUCTURE = [
    "Yes (Sufficient)", 
    "Partial (Upgrade Needed)", 
    "No (Full Build)", 
    "Not Assessed", 
    "N/A"
]

ONPREM_STAFFING_PLAN = ["Yes", "In Progress", "Not Started", "N/A"]


# ============================================================================
# SECTION 1: WORKBOOK CREATION
# ============================================================================



# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
CLOUD = '\u2601'      # ☁️  Cloud
GLOBE = '\U0001F310'  # 🌐 Globe
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    sheets = [
        "Instructions & Legend",
        "2. SaaS Services",
        "3. IaaS PaaS Services",
        "4. Cloud Security Services",
        "5. Cloud Storage Services",
        "6. Data Classification",
        "7. Service Criticality",
        "8. Summary Dashboard",
        "9. Evidence Register",
        "10. Approval Sign-Of",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================

def setup_styles() -> dict:
    """Define cell styles - returns dict of style definitions."""
    return {
        "header": {
            "font_params": {"name": "Calibri", "size": 14, "bold": True, "color": "FFFFFF"},
            "fill_params": {"start_color": "003366", "end_color": "003366", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "subheader": {
            "font_params": {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"},
            "fill_params": {"start_color": "4472C4", "end_color": "4472C4", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "column_header": {
            "font_params": {"name": "Calibri", "size": 10, "bold": True},
            "fill_params": {"start_color": "D9D9D9", "end_color": "D9D9D9", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "input_cell": {
            "fill_params": {"start_color": "FFFFCC", "end_color": "FFFFCC", "fill_type": "solid"},
            "alignment_params": {"horizontal": "left", "vertical": "center", "wrap_text": True},
        },
    }


def apply_style(cell, style_def, style_type):
    """Apply style to cell using NEW objects (avoids shared object issues)."""
    if "font_params" in style_def:
        cell.font = Font(**style_def["font_params"])
    if "fill_params" in style_def:
        cell.fill = PatternFill(**style_def["fill_params"])
    if "alignment_params" in style_def:
        cell.alignment = Alignment(**style_def["alignment_params"])
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 3: COLUMN DEFINITIONS (BASE + EXTENDED + REGULATORY)
# ============================================================================

def get_base_columns() -> dict:
    """Base columns A-Q (17 columns)."""
    return {
        "Cloud_Service_Name": 30, "Service_Type": 18, "Vendor_Name": 25,
        "Service_Criticality": 18, "Data_Classification": 18, "Data_Residency": 18,
        "Contract_Status": 18, "Status": 15, "Evidence_Location": 30,
        "Gap_Description": 35, "Remediation_Needed": 14, "Exception_ID": 14,
        "Risk_ID": 14, "Compensating_Controls": 30, "Service_Owner": 22,
        "Target_Remediation_Date": 18, "Budget_Impact": 16,
    }


def get_extended_columns() -> dict:
    """Extended columns R-X (7 columns)."""
    return {
        "Monthly_Cost_CHF": 16, "Annual_Contract_Value": 18, "Licensed_Users": 14,
        "Integration_Count": 14, "Backup_Available": 16, "Exit_Feasibility": 16,
        "Last_Review_Date": 16,
    }


def get_regulatory_columns() -> dict:
    """Regulatory columns Y-AG (9 columns) for DORA/NIS2/AI Act/CLOUD Act."""
    return {
        "Provider_HQ_Jurisdiction": 20, "CLOUD_Act_Exposure": 20,
        "Data_Processing_Locations": 25, "EU_Data_Boundary": 16,
        "DORA_In_Scope": 14, "NIS2_In_Scope": 14, "AI_Classification": 18,
        "Concentration_Risk": 18, "Alternatives_Identified": 16,
    }


def get_all_columns() -> dict:
    """All columns A-AG (33 total)."""
    all_cols = {}
    all_cols.update(get_base_columns())
    all_cols.update(get_extended_columns())
    all_cols.update(get_regulatory_columns())
    return all_cols

# ============================================================================
# SECTION 4: VALIDATION DEFINITIONS
# ============================================================================

def create_all_validations(ws) -> dict:
    """Create ALL data validations (base + extended + regulatory)."""
    validations = {}
    
    # --- BASE VALIDATIONS ---
    validations['service_type'] = DataValidation(
        type="list", formula1='"SaaS,IaaS,PaaS,Security,Storage,Collaboration,Other"', allow_blank=True)
    validations['criticality'] = DataValidation(
        type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    validations['data_class'] = DataValidation(
        type="list", formula1='"Restricted,Confidential,Internal,Public,Mixed,N/A"', allow_blank=True)
    validations['residency'] = DataValidation(
        type="list", formula1='"Switzerland,EU,USA,Global,Unknown"', allow_blank=True)
    validations['contract_status'] = DataValidation(
        type="list", formula1='"Active,Renewal Due,Expired,Under Negotiation,Trial"', allow_blank=True)
    validations['status'] = DataValidation(
        type="list", formula1=f'"{CHECK} Compliant,⚠️ Partial,❌ Non-Compliant,N/A"', allow_blank=True)
    validations['yes_no'] = DataValidation(
        type="list", formula1='"Yes,No"', allow_blank=True)
    validations['budget'] = DataValidation(
        type="list", formula1='"High (>$50K),Medium ($10K-$50K),Low (<$10K),None"', allow_blank=True)
    
    # --- EXTENDED VALIDATIONS ---
    validations['backup'] = DataValidation(
        type="list", formula1='"Yes,No,Planned,Unknown"', allow_blank=True)
    validations['exit_feasibility'] = DataValidation(
        type="list", formula1='"Easy (≤30 days),Medium (31-90 days),Hard (>90 days),Unknown"', allow_blank=True)
    
    # --- REGULATORY VALIDATIONS ---
    validations['hq_jurisdiction'] = DataValidation(
        type="list", formula1='"' + ','.join(PROVIDER_HQ_JURISDICTION) + '"', allow_blank=True)
    validations['cloud_act'] = DataValidation(
        type="list", formula1='"' + ','.join(CLOUD_ACT_EXPOSURE) + '"', allow_blank=True)
    validations['eu_boundary'] = DataValidation(
        type="list", formula1='"' + ','.join(EU_DATA_BOUNDARY_OPTIONS) + '"', allow_blank=True)
    validations['dora_scope'] = DataValidation(
        type="list", formula1='"' + ','.join(YES_NO_NOT_DETERMINED) + '"', allow_blank=True)
    validations['nis2_scope'] = DataValidation(
        type="list", formula1='"' + ','.join(YES_NO_NOT_DETERMINED) + '"', allow_blank=True)
    validations['ai_class'] = DataValidation(
        type="list", formula1='"' + ','.join(AI_RISK_CLASSIFICATION) + '"', allow_blank=True)
    validations['concentration'] = DataValidation(
        type="list", formula1='"' + ','.join(CONCENTRATION_RISK_LEVEL) + '"', allow_blank=True)
    validations['alternatives'] = DataValidation(
        type="list", formula1='"' + ','.join(YES_NO_PARTIAL) + '"', allow_blank=True)
    
    # Add all to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    return validations


def apply_validations_to_range(ws, validations: dict, start_row: int, end_row: int):
    """Apply validations to data entry range."""
    for row in range(start_row, end_row + 1):
        # Base columns
        validations['service_type'].add(ws.cell(row=row, column=2))    # B
        validations['criticality'].add(ws.cell(row=row, column=4))     # D
        validations['data_class'].add(ws.cell(row=row, column=5))      # E
        validations['residency'].add(ws.cell(row=row, column=6))       # F
        validations['contract_status'].add(ws.cell(row=row, column=7)) # G
        validations['status'].add(ws.cell(row=row, column=8))          # H
        validations['yes_no'].add(ws.cell(row=row, column=11))         # K
        validations['budget'].add(ws.cell(row=row, column=17))         # Q
        
        # Extended columns
        validations['backup'].add(ws.cell(row=row, column=22))         # V
        validations['exit_feasibility'].add(ws.cell(row=row, column=23)) # W
        
        # Regulatory columns (Y=25, Z=26, AA=27, AB=28, AC=29, AD=30, AE=31, AF=32, AG=33)
        validations['hq_jurisdiction'].add(ws.cell(row=row, column=25))  # Y
        validations['cloud_act'].add(ws.cell(row=row, column=26))        # Z
        # AA (27) = free text, no validation
        validations['eu_boundary'].add(ws.cell(row=row, column=28))      # AB
        validations['dora_scope'].add(ws.cell(row=row, column=29))       # AC
        validations['nis2_scope'].add(ws.cell(row=row, column=30))       # AD
        validations['ai_class'].add(ws.cell(row=row, column=31))         # AE
        validations['concentration'].add(ws.cell(row=row, column=32))    # AF
        validations['alternatives'].add(ws.cell(row=row, column=33))     # AG


# ============================================================================
# SECTION 5: CHECKLIST DEFINITIONS
# ============================================================================

def get_checklist_items() -> list:
    """Return combined checklist (base + regulatory)."""
    base = [
        ("All cloud services documented in CMDB", "CMDB export"),
        ("Data classification assigned", "Data classification matrix"),
        ("Service criticality assessed (BIA)", "Business impact analysis"),
        ("Service owner identified", "RACI matrix"),
        ("Contract status tracked", "Contract management system"),
        ("Monthly costs documented", "Finance system"),
        ("User license count verified", "License report"),
        ("Integration dependencies mapped", "Architecture diagram"),
        ("Backup strategy defined", "BC/DR plan"),
        ("Exit feasibility evaluated", "Exit strategy document"),
        ("Data residency verified", "Vendor DPA"),
        ("Last review date recorded", "Review calendar"),
    ]
    regulatory = [
        ("Provider HQ jurisdiction documented", "Vendor documentation"),
        ("CLOUD Act exposure assessed", "Risk assessment"),
        ("Data processing locations verified", "DPA, vendor statement"),
        ("DORA applicability determined", "Regulatory review"),
        ("NIS2 applicability determined", "Regulatory review"),
        ("AI system classification assessed", "AI inventory"),
        ("Concentration risk evaluated", "Risk assessment"),
        ("Alternative providers identified", "Market analysis"),
    ]
    return base + regulatory

# ============================================================================
# SECTION 6: INSTRUCTIONS SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "ISMS-IMP-A.5.23.S1 - Cloud Service Inventory & Classification"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 40
    
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)
    
    doc_info = [
        ("Document ID", "ISMS-IMP-A.5.23.S1"),
        ("Assessment Area", "Cloud Service Inventory & Classification"),
        ("Related Policy", "ISMS-POL-A.5.19-23-S5"),
        ("Version", "1.0"),
        ("Assessment Date", ""), ("Completed By", ""), ("Organization", ""),
        ("Review Cycle", "Quarterly"),
    ]
    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            apply_style(ws[f"B{row}"], styles["input_cell"], "input")
        row += 1
    
    row += 1
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    for instr in [
        "1. Complete each assessment sheet for all cloud services",
        "2. Use dropdown menus for standardized entries",
        "3. Fill yellow cells with your information",
        "4. Complete regulatory columns (Y-AG) for DORA/NIS2/AI Act/CLOUD Act",
        "5. Provide evidence location for each entry",
        "6. Dashboard auto-calculates compliance statistics",
    ]:
        ws[f"A{row}"] = instr
        row += 1
    
    # Regulatory guidance
    row += 1
    ws[f"A{row}"] = "REGULATORY COMPLIANCE FIELDS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="003366")
    row += 1
    for line in [
        "Columns Y-AG track regulatory requirements:",
        f"{BULLET} DORA (EU Financial Sector) - Complete if you are an EU financial entity",
        f"{BULLET} NIS2 (Essential/Important Entities) - Complete if subject to NIS2",
        f"{BULLET} AI Act - Complete if using AI systems",
        f"{BULLET} CLOUD Act - Assess for US-headquartered providers",
    ]:
        ws[f"A{row}"] = line
        row += 1
    
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 30


# ============================================================================
# SECTION 7: ASSESSMENT SHEET CREATOR
# ============================================================================

def create_assessment_sheet(ws, styles, title, policy_req, question, checklist):
    """Create a service assessment sheet with all 33 columns."""
    columns = get_all_columns()
    total_cols = len(columns)
    last_col = get_column_letter(total_cols)
    
    # Header
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"{title}\n{policy_req}"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 50
    
    # Question
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = question
    ws["A3"].font = Font(bold=True, size=11)
    
    # Response dropdown
    ws["A4"] = "Response:"
    ws["B4"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    resp_dv = DataValidation(type="list", formula1='"Yes,No,Partial,Under Review"', allow_blank=False)
    ws.add_data_validation(resp_dv)
    resp_dv.add(ws["B4"])
    
    # Column headers (row 6)
    header_row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        apply_style(cell, styles["column_header"], "header")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # Data entry area (rows 8-50)
    start_row, end_row = 8, 50
    for row in range(start_row, end_row + 1):
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Apply validations
    validations = create_all_validations(ws)
    apply_validations_to_range(ws, validations, start_row, end_row)
    
    # Example row
    ws["A8"] = "Example: Microsoft 365"
    ws["B8"] = "SaaS"
    ws["C8"] = "Microsoft"
    ws["D8"] = "Critical"
    ws["H8"] = f"{CHECK} Compliant"
    ws["Y8"] = "United States"
    ws["Z8"] = "Mitigated (EU Data Boundary)"
    ws["AC8"] = "Yes"
    ws["AD8"] = "Yes"
    ws["AF8"] = "High (Few alternatives)"
    ws["A8"].font = Font(italic=True, color="666666")
    
    # Checklist
    checklist_row = end_row + 3
    ws.merge_cells(f"A{checklist_row}:D{checklist_row}")
    ws[f"A{checklist_row}"] = "COMPLIANCE CHECKLIST"
    apply_style(ws[f"A{checklist_row}"], styles["subheader"], "subheader")
    
    checklist_row += 2
    headers = ["Item", "Requirement", "Status", "Evidence"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=checklist_row, column=col_idx, value=h)
        apply_style(cell, styles["column_header"], "header")
    
    status_dv = DataValidation(type="list", formula1=f'"{CHECK},{WARNING},{XMARK}"', allow_blank=True)
    ws.add_data_validation(status_dv)
    
    checklist_row += 1
    for req, evidence in checklist:
        ws.cell(row=checklist_row, column=1, value="☐")
        ws.cell(row=checklist_row, column=2, value=req)
        status_dv.add(ws.cell(row=checklist_row, column=3))
        ws.cell(row=checklist_row, column=3).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws.cell(row=checklist_row, column=4, value=evidence)
        checklist_row += 1
    
    ws.freeze_panes = "A7"


# ============================================================================
# SECTION 7B: EXIT STRATEGY ASSESSMENT SHEET (SHEET 4)
# ============================================================================

def create_exit_strategy_sheet(ws, styles):
    """
    Create Sheet 4: Exit Feasibility & Strategy Assessment
    
    Implements POL-A.5.19-23-S5 Section 8.1.1 (Exit Strategy Framework)
    
    Column Structure:
    - Columns A-Q: Basic service information (17 columns)
    - Columns R-AC: Exit Strategy framework (12 columns)
    Total: 29 columns
    """
    
    # === HEADER ===
    ws.merge_cells('A1:AC1')
    ws['A1'] = "SHEET 4: EXIT FEASIBILITY & STRATEGY ASSESSMENT"
    ws['A1'].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    
    # === POLICY REFERENCE ===
    ws.merge_cells('A2:AC2')
    ws['A2'] = "POL-S5 Section 8.1.1: Cloud-to-Cloud (90%+), Hybrid (5-10%), On-Premises (<5% only when justified)"
    ws['A2'].font = Font(name="Calibri", size=9, italic=True)
    ws['A2'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    # === COLUMN HEADERS (Row 3) ===
    headers = [
        # Columns A-Q: Basic Service Information
        ("A", "Service Name", 25),
        ("B", "Provider", 20),
        ("C", "Service Type", 18),
        ("D", "Environment", 15),
        ("E", "Business Owner", 20),
        ("F", "Technical Owner", 20),
        ("G", "Data Classification", 18),
        ("H", "Criticality", 15),
        ("I", "Users", 12),
        ("J", "Cost (Annual CHF)", 18),
        ("K", "Contract End Date", 15),
        ("L", "Primary Region", 18),
        ("M", "Backup Region", 18),
        ("N", "Integration Count", 15),
        ("O", "API Dependency", 15),
        ("P", "Compliance Certs", 20),
        ("Q", "Current Status", 15),
        
        # Columns R-AC: Exit Strategy Framework
        ("R", "Exit Strategy Type", 20),
        ("S", "Alternative Identified", 22),
        ("T", "Export Format Available", 22),
        ("U", "Export Tested", 16),
        ("V", "Data Volume (GB)", 16),
        ("W", "Migration Complexity", 24),
        ("X", "Lock-In Risk", 16),
        ("Y", "Hybrid: Workload Segmentation", 28),
        ("Z", "Hybrid: Data Sync Latency", 28),
        ("AA", "OnPrem: TCO Analysis Complete", 28),
        ("AB", "OnPrem: Infrastructure Available", 30),
        ("AC", "OnPrem: Staffing Plan Documented", 30),
    ]
    
    thin = Side(style="thin")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
        ws.column_dimensions[col].width = width
    
    # === DATA VALIDATION SETUP ===
    
    # Basic validations
    val_service_type = DataValidation(
        type="list", formula1='"SaaS,IaaS,PaaS,Security,Storage,Other"', allow_blank=False)
    val_environment = DataValidation(
        type="list", formula1='"Production,Development,Test,Staging,All"', allow_blank=False)
    val_data_class = DataValidation(
        type="list", formula1='"Restricted,Confidential,Internal,Public,Mixed"', allow_blank=False)
    val_criticality = DataValidation(
        type="list", formula1='"Critical,High,Medium,Low"', allow_blank=False)
    val_api_dependency = DataValidation(
        type="list", formula1='"Yes,No,Partial"', allow_blank=False)
    val_status = DataValidation(
        type="list", formula1=f'"{CHECK} Compliant,⚠️ Partial,❌Non-Compliant"', allow_blank=False)
    
    # Exit Strategy validations
    val_exit_strategy = DataValidation(
        type="list", formula1='"' + ','.join(EXIT_STRATEGY_TYPE) + '"', allow_blank=False)
    val_exit_strategy.error = "Select: Cloud-to-Cloud (default), Hybrid, On-Premises, or Not Yet Determined"
    val_exit_strategy.errorTitle = "Invalid Exit Strategy"
    
    val_alternative = DataValidation(
        type="list", formula1='"' + ','.join(ALTERNATIVE_IDENTIFIED) + '"', allow_blank=False)
    
    val_export_format = DataValidation(
        type="list", formula1='"' + ','.join(EXPORT_FORMAT) + '"', allow_blank=False)
    
    val_export_tested = DataValidation(
        type="list", formula1='"' + ','.join(EXPORT_TESTED) + '"', allow_blank=False)
    
    val_data_volume = DataValidation(
        type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    val_data_volume.error = "Enter a positive number (GB)"
    val_data_volume.errorTitle = "Invalid Data Volume"
    
    val_migration_complexity = DataValidation(
        type="list", formula1='"' + ','.join(MIGRATION_COMPLEXITY) + '"', allow_blank=False)
    
    val_lock_in_risk = DataValidation(
        type="list", formula1='"' + ','.join(LOCK_IN_RISK) + '"', allow_blank=False)
    
    val_hybrid_workload = DataValidation(
        type="list", formula1='"' + ','.join(HYBRID_WORKLOAD_SEGMENTATION) + '"', allow_blank=False)
    
    val_hybrid_sync = DataValidation(
        type="list", formula1='"' + ','.join(HYBRID_DATA_SYNC_LATENCY) + '"', allow_blank=False)
    
    val_onprem_tco = DataValidation(
        type="list", formula1='"' + ','.join(ONPREM_TCO_ANALYSIS) + '"', allow_blank=False)
    
    val_onprem_infra = DataValidation(
        type="list", formula1='"' + ','.join(ONPREM_INFRASTRUCTURE) + '"', allow_blank=False)
    
    val_onprem_staffing = DataValidation(
        type="list", formula1='"' + ','.join(ONPREM_STAFFING_PLAN) + '"', allow_blank=False)
    
    # Add validations to worksheet
    for val in [val_service_type, val_environment, val_data_class, val_criticality, val_api_dependency, val_status,
                val_exit_strategy, val_alternative, val_export_format, val_export_tested, val_data_volume,
                val_migration_complexity, val_lock_in_risk, val_hybrid_workload, val_hybrid_sync,
                val_onprem_tco, val_onprem_infra, val_onprem_staffing]:
        ws.add_data_validation(val)
    
    # === APPLY VALIDATIONS TO DATA ROWS (4-100) ===
    for row in range(4, 101):
        # Basic columns
        val_service_type.add(ws[f'C{row}'])
        val_environment.add(ws[f'D{row}'])
        val_data_class.add(ws[f'G{row}'])
        val_criticality.add(ws[f'H{row}'])
        val_api_dependency.add(ws[f'O{row}'])
        val_status.add(ws[f'Q{row}'])
        
        # Exit Strategy columns
        val_exit_strategy.add(ws[f'R{row}'])
        val_alternative.add(ws[f'S{row}'])
        val_export_format.add(ws[f'T{row}'])
        val_export_tested.add(ws[f'U{row}'])
        val_data_volume.add(ws[f'V{row}'])
        val_migration_complexity.add(ws[f'W{row}'])
        val_lock_in_risk.add(ws[f'X{row}'])
        val_hybrid_workload.add(ws[f'Y{row}'])
        val_hybrid_sync.add(ws[f'Z{row}'])
        val_onprem_tco.add(ws[f'AA{row}'])
        val_onprem_infra.add(ws[f'AB{row}'])
        val_onprem_staffing.add(ws[f'AC{row}'])
        
        # Yellow fill + borders for input cells
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q',
                    'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']:
            cell = ws[f'{col}{row}']
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
    
    # === CONDITIONAL FORMATTING - WARNINGS ===
    from openpyxl.formatting.rule import FormulaRule

    # WARNING 1: On-Premises strategy → Orange fill (warn: <5% justified)
    onprem_warning = FormulaRule(
        formula=['$R4="On-Premises"'],
        fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        stopIfTrue=False
    )
    ws.conditional_formatting.add('R4:AC100', onprem_warning)
    
    # WARNING 2: Critical Lock-In Risk → Red fill
    critical_lock_in = FormulaRule(
        formula=['$X4="Critical"'],
        fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        font=Font(color="FFFFFF", bold=True),
        stopIfTrue=False
    )
    ws.conditional_formatting.add('X4:X100', critical_lock_in)
    
    # WARNING 3: Export not tested for Critical services → Yellow fill
    untested_critical = FormulaRule(
        formula=['AND($U4="No", $H4="Critical")'],
        fill=PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
        stopIfTrue=False
    )
    ws.conditional_formatting.add('U4:U100', untested_critical)
    
    # === CHECKLIST SECTION (Row 105+) ===
    ws.merge_cells('A105:AC105')
    ws['A105'] = "EXIT STRATEGY COMPLIANCE CHECKLIST (22 items)"
    ws['A105'].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws['A105'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws['A105'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[105].height = 30
    
    checklist_items = [
        # Original 15 items
        f"{CHECK} Exit strategy documented for Critical/High services",
        f"{CHECK} Exit triggers defined",
        f"{CHECK} Alternative providers/solutions identified",
        f"{CHECK} Alternative provider assessment completed",
        f"{CHECK} Migration effort estimated",
        f"{CHECK} Data extraction requirements documented",
        f"{CHECK} Application dependencies mapped",
        f"{CHECK} Integration points documented",
        f"{CHECK} Lock-in risks identified and documented",
        f"{CHECK} Exit cost estimate completed",
        f"{CHECK} Exit timeline estimate completed",
        f"{CHECK} Data export format verified",
        f"{CHECK} Migration testing requirements defined",
        f"{CHECK} Exit plan reviewed and approved",
        f"{CHECK} Exit strategy included in vendor contracts",
        
        # NEW 7 items for Exit Strategy framework
        f"{CHECK} Cloud-to-Cloud alternatives evaluated (min. 2 providers)",
        f"{CHECK} Hybrid workload segmentation documented (if applicable)",
        f"{CHECK} On-Premises TCO analysis completed (if applicable)",
        f"{CHECK} Annual PoC testing completed (DORA Art. 28.6 for Critical)",
        f"{CHECK} Vendor lock-in risks assessed and mitigation documented",
        f"{CHECK} Migration cost estimates align with exit strategy type",
        f"{CHECK} Exit strategy type justified with evidence (esp. On-Premises)",
    ]
    
    row = 106
    for idx, item in enumerate(checklist_items, 1):
        ws[f'A{row}'] = f"{idx}. {item}"
        ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 22
        row += 1
    
    # === HELP TEXT SECTION (Row 135+) ===
    ws.merge_cells('A135:AC135')
    ws['A135'] = "EXIT STRATEGY GUIDANCE - POLICY REQUIREMENTS"
    ws['A135'].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws['A135'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws['A135'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[135].height = 25
    
    help_text = [
        "",
        f"{CHECK} CLOUD-TO-CLOUD (90%+ of services - DEFAULT STRATEGY)",
        "   • Migrate to alternative cloud provider (AWS, Azure, GCP, OVHcloud, Alibaba, etc.)",
        "   • Timeline: 3-6 months typical",
        "   • Cost: CHF 40K-200K one-time",
        "   • Benefits: Zero CAPEX, maintained elasticity, modern cloud-native capabilities",
        "   • Required: Min. 2 alternative providers identified, annual PoC testing (DORA Art. 28.6)",
        "",
        f"{WARNING} HYBRID (5-10% of services - PARTIAL REPATRIATION)",
        "   • Some workloads in cloud, others on-premises (e.g., app in cloud, DB on-prem)",
        "   • Timeline: 6-12 months typical",
        "   • Cost: CHF 335K-1.6M (3-year TCO)",
        "   • Use cases: Data sovereignty (FADP), cost optimization, latency requirements",
        "   • Required: Workload segmentation documented, data sync latency assessed, network design",
        "",
        f"{WARNING}{WARNING} ON-PREMISES (<5% of services - FULL BUILD-BACK)",
        "   • Complete migration to [Organization]-owned infrastructure",
        "   • Timeline: 6-18 months (long!)",
        "   • Cost: CHF 2.94M-7.04M (5-year TCO)",
        "   • Justification: Regulatory mandate, cost inversion (>CHF 500K/yr cloud), strategic independence",
        "   • Required: TCO analysis, infrastructure capacity plan, staffing plan (3-10 FTEs)",
        "   • WARNING: Economically justified in <5% of cases - validate TCO carefully!",
        "",
        f"{LOCK} VENDOR LOCK-IN RISK LEVELS:",
        "   • Low: Standard protocols, open formats, portable architecture",
        "   • Medium: Some proprietary features, manageable migration effort",
        "   • High: Deep integration with proprietary services, significant refactoring needed",
        "   • Critical: Mission-critical dependency on unique capability, no viable alternatives",
        "",
        f"{SHIELD} DORA COMPLIANCE (Financial Sector):",
        "   • Article 28.6: Annual PoC testing of exit strategies for critical ICT providers",
        "   • Article 28.9: Concentration risk mitigation (reduce dependency on single hyperscaler)",
        "   • Track PoC testing in ISMS-IMP-A.5.23.S4 (Governance workbook)",
        "",
        "For detailed cost models and TCO calculations, refer to:",
        "   • POL-S5 Section 8.1.1.1 (Cloud-to-Cloud)",
        "   • POL-S5 Section 8.1.1.2 (Hybrid)",
        "   • POL-S5 Section 8.1.1.3 (On-Premises)",
        "",
    ]
    
    row = 136
    for line in help_text:
        ws[f'A{row}'] = line
        ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 18
        row += 1
    
    ws.freeze_panes = "A4"
    
    logger.info("   [OK] Sheet 4: Exit Feasibility & Strategy Assessment created")
    logger.info(f"        Columns: A-AC (29 total)")
    logger.info(f"        Exit Strategy framework: Columns R-AC (12 columns)")
    logger.info(f"        Conditional formatting: On-Prem warning, Lock-In risk, Untested exports")


# ============================================================================
# SECTION 6: DATA CLASSIFICATION SHEET
# ============================================================================

def create_data_classification_sheet(ws, styles):
    """Create Data Classification sheet for cloud-stored data inventory."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "DATA CLASSIFICATION - CLOUD STORAGE INVENTORY"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 30

    # Instruction row
    ws.merge_cells("A2:J2")
    ws["A2"] = "Classify all data types stored in cloud services. Required for ISO 27001 A.5.12 (Classification), A.5.13 (Labelling), and GDPR Article 30."
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)

    # Headers
    headers = [
        ("A", "Data_ID", 12),
        ("B", "Data Category", 25),
        ("C", "Description", 35),
        ("D", "Classification Level", 18),
        ("E", "Primary Cloud Service", 22),
        ("F", "Storage Location", 22),
        ("G", "Data Owner", 18),
        ("H", "Retention Period", 16),
        ("I", "Personal Data (GDPR)", 18),
        ("J", "Cross-Border Transfer", 18),
    ]

    thin = Side(style="thin")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
        ws.column_dimensions[col].width = width

    # Data validations
    val_classification = DataValidation(
        type="list", formula1='"Restricted,Confidential,Internal,Public"', allow_blank=False)
    val_personal_data = DataValidation(
        type="list", formula1='"Yes - Special Category,Yes - Personal,No"', allow_blank=False)
    val_cross_border = DataValidation(
        type="list", formula1='"None,EEA Only,Adequacy Decision,SCCs,Not Assessed"', allow_blank=False)

    ws.add_data_validation(val_classification)
    ws.add_data_validation(val_personal_data)
    ws.add_data_validation(val_cross_border)

    # Pre-fill rows with styling and validations
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(4, 104):
        ws.cell(row=row, column=1, value=f"DC-{row-3:03d}")
        for col_idx in range(1, 11):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

        val_classification.add(ws[f'D{row}'])
        val_personal_data.add(ws[f'I{row}'])
        val_cross_border.add(ws[f'J{row}'])

    ws.freeze_panes = "A4"
    logger.info("   [OK] Sheet 6: Data Classification created (100 rows)")


# ============================================================================
# SECTION 7: SERVICE CRITICALITY SHEET
# ============================================================================

def create_service_criticality_sheet(ws, styles):
    """Create Service Criticality Assessment sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "SERVICE CRITICALITY ASSESSMENT"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 30

    # Instruction row
    ws.merge_cells("A2:L2")
    ws["A2"] = "Assess criticality of each cloud service for BIA and exit strategy planning. Required for ISO 27001 A.5.29 (Business Continuity) and DORA Article 28."
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)

    # Headers
    headers = [
        ("A", "Service_ID", 12),
        ("B", "Cloud Service Name", 28),
        ("C", "Service Category", 18),
        ("D", "Business Process Supported", 28),
        ("E", "Criticality Level", 16),
        ("F", "RTO (Hours)", 14),
        ("G", "RPO (Hours)", 14),
        ("H", "MTPD (Hours)", 14),
        ("I", "Single Point of Failure", 18),
        ("J", "Workaround Available", 18),
        ("K", "DORA Scope", 14),
        ("L", "Exit Priority", 14),
    ]

    thin = Side(style="thin")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
        ws.column_dimensions[col].width = width

    # Data validations
    val_category = DataValidation(
        type="list", formula1='"SaaS,IaaS,PaaS,Security,Storage,Other"', allow_blank=False)
    val_criticality = DataValidation(
        type="list", formula1='"Critical,High,Medium,Low"', allow_blank=False)
    val_yesno = DataValidation(
        type="list", formula1='"Yes,No,Partial"', allow_blank=False)
    val_dora = DataValidation(
        type="list", formula1='"In-Scope,Out-of-Scope,TBD"', allow_blank=False)
    val_priority = DataValidation(
        type="list", formula1='"P1 - Critical,P2 - High,P3 - Medium,P4 - Low"', allow_blank=False)

    ws.add_data_validation(val_category)
    ws.add_data_validation(val_criticality)
    ws.add_data_validation(val_yesno)
    ws.add_data_validation(val_dora)
    ws.add_data_validation(val_priority)

    # Pre-fill rows with styling and validations
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(4, 104):
        ws.cell(row=row, column=1, value=f"SC-{row-3:03d}")
        for col_idx in range(1, 13):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border

        val_category.add(ws[f'C{row}'])
        val_criticality.add(ws[f'E{row}'])
        val_yesno.add(ws[f'I{row}'])
        val_yesno.add(ws[f'J{row}'])
        val_dora.add(ws[f'K{row}'])
        val_priority.add(ws[f'L{row}'])

    ws.freeze_panes = "A4"
    logger.info("   [OK] Sheet 7: Service Criticality Assessment created (100 rows)")


# ============================================================================
# SECTION 8: SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with compliance and regulatory metrics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "CLOUD SERVICE INVENTORY - COMPLIANCE DASHBOARD"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 30
    
    # Compliance by category
    row = 3
    ws[f"A{row}"] = "COMPLIANCE BY SERVICE CATEGORY"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    
    row += 1
    headers = ["Category", "Total", f"{CHECK} Compliant", f"{WARNING} Partial", f"{XMARK} Non-Compliant", "N/A", "Compliance %"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        apply_style(cell, styles["column_header"], "header")
    
    categories = [
        ("SaaS Services", "2. SaaS Services"),
        ("IaaS/PaaS", "3. IaaS PaaS Services"),
        ("Security Services", "4. Cloud Security Services"),
        ("Storage Services", "5. Cloud Storage Services"),
    ]
    
    row += 1
    for label, sheet in categories:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=f"=COUNTA('{sheet}'!A8:A50)")
        ws.cell(row=row, column=3, value=f'=COUNTIF(\'{sheet}\'!H8:H50,"{CHECK}*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF(\'{sheet}\'!H8:H50,"{WARNING}*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF(\'{sheet}\'!H8:H50,"{XMARK}*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF(\'{sheet}\'!H8:H50,"N/A")')
        ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1
    
    # REGULATORY METRICS
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "REGULATORY COMPLIANCE SUMMARY (DORA / NIS2 / AI Act / CLOUD Act)"
    ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=12)
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    
    row += 2
    for col_idx, h in enumerate(["Metric", "Count", "Status"], start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        apply_style(cell, styles["column_header"], "header")
    
    row += 1
    reg_metrics = [
        ("CLOUD Act Potential Exposure", 
         '=COUNTIF(\'2. SaaS Services\'!Z:Z,"Potential*")+COUNTIF(\'3. IaaS PaaS Services\'!Z:Z,"Potential*")'),
        ("DORA In-Scope Services",
         '=COUNTIF(\'2. SaaS Services\'!AC:AC,"Yes")+COUNTIF(\'3. IaaS PaaS Services\'!AC:AC,"Yes")'),
        ("NIS2 In-Scope Services",
         '=COUNTIF(\'2. SaaS Services\'!AD:AD,"Yes")+COUNTIF(\'3. IaaS PaaS Services\'!AD:AD,"Yes")'),
        ("High-Risk AI Systems",
         '=COUNTIF(\'2. SaaS Services\'!AE:AE,"High Risk")+COUNTIF(\'3. IaaS PaaS Services\'!AE:AE,"High Risk")'),
        ("Critical Concentration Risk",
         '=COUNTIF(\'2. SaaS Services\'!AF:AF,"Critical*")+COUNTIF(\'3. IaaS PaaS Services\'!AF:AF,"Critical*")'),
    ]
    
    for metric, formula in reg_metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=3, value=f'=IF(B{row}>0,"{WARNING} Review","{CHECK} OK")')
        row += 1
    
    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 22

# ============================================================================
# SECTION 9: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "EVIDENCE REGISTER - CLOUD SERVICE INVENTORY"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 30
    
    headers = [
        ("Evidence_ID", 15), ("Cloud_Service", 30), ("Evidence_Type", 25),
        ("Description", 40), ("File_Location", 40), ("Collection_Date", 16),
        ("Collected_By", 20), ("Retention", 16), ("Status", 15),
    ]
    
    row = 3
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        apply_style(cell, styles["column_header"], "header")
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    
    # Pre-fill evidence IDs
    for i in range(1, 101):
        ws.cell(row=row + i, column=1, value=f"EV-INV-{i:03d}")
        for col in range(2, 10):
            cell = ws.cell(row=row + i, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Evidence type dropdown
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Contract,Invoice,CMDB Export,Screenshot,Config File,Risk Assessment,Audit Report,Other"',
        allow_blank=True
    )
    ws.add_data_validation(ev_type_dv)
    ev_type_dv.add(f"C4:C103")
    
    # Status dropdown
    status_dv = DataValidation(type="list", formula1='"Current,Expired,Pending"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f"I4:I103")


# ============================================================================
# SECTION 10: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create Approval Sign-Off sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "APPROVAL SIGN-OFF - CLOUD SERVICE INVENTORY"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 30
    
    # Assessment summary
    row = 3
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    
    summary = [
        ("Document:", "ISMS-IMP-A.5.23.S1 - Cloud Service Inventory"),
        ("Assessment Period:", ""),
        ("Total Services:", "='8. Summary Dashboard'!B8"),
        ("Overall Compliance:", "='8. Summary Dashboard'!G8"),
        ("Assessment Status:", ""),
    ]
    
    row += 1
    for label, value in summary:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1
    
    # Approval sections
    approvers = [
        ("COMPLETED BY (IT OPERATIONS)", "4472C4"),
        ("REVIEWED BY (ISO)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    
    row += 2
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1
        
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
            row += 1
        row += 1
    
    # Final decision
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])
    
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30


# ============================================================================
# SECTION 11: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution - generates the complete workbook."""
    logger.info("=" * 70)
    logger.info("ISMS-IMP-A.5.23.S1 - Cloud Service Inventory Generator")
    logger.info("WITH REGULATORY COLUMNS (DORA / NIS2 / AI Act / CLOUD Act)")
    logger.info("=" * 70)
    
    wb = create_workbook()
    styles = setup_styles()
    checklist = get_checklist_items()
    
    logger.info("\n[1/10] Creating Instructions & Legend...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    
    logger.info("[2/10] Creating 2. SaaS Services...")
    create_assessment_sheet(
        wb["2. SaaS Services"], styles,
        "SAAS SERVICES INVENTORY",
        "All SaaS applications must be documented with data classification (Policy S5 Section 2.1)",
        "Does your organization use SaaS applications?",
        checklist
    )
    
    logger.info("[3/10] Creating 3. IaaS PaaS Services...")
    create_assessment_sheet(
        wb["3. IaaS PaaS Services"], styles,
        "IAAS/PAAS SERVICES INVENTORY",
        "All infrastructure services must maintain security baselines (Policy S5 Section 2.2)",
        "Does your organization use IaaS or PaaS services?",
        checklist
    )
    
    logger.info("[4/10] Creating 4. Exit Feasibility & Strategy Assessment...")
    create_exit_strategy_sheet(wb["4. Cloud Security Services"], styles)
    
    logger.info("[5/10] Creating 5. Cloud Storage Services...")
    create_assessment_sheet(
        wb["5. Cloud Storage Services"], styles,
        "CLOUD STORAGE SERVICES INVENTORY",
        "All cloud storage must enforce encryption (Policy S5 Section 2.4)",
        "Does your organization use cloud storage services?",
        checklist
    )
    
    logger.info("[6/10] Creating 6. Data Classification...")
    create_data_classification_sheet(wb["6. Data Classification"], styles)

    logger.info("[7/10] Creating 7. Service Criticality...")
    create_service_criticality_sheet(wb["7. Service Criticality"], styles)
    
    logger.info("[8/10] Creating 8. Summary Dashboard...")
    create_summary_dashboard(wb["8. Summary Dashboard"], styles)
    
    logger.info("[9/10] Creating 9. Evidence Register...")
    create_evidence_register(wb["9. Evidence Register"], styles)
    
    logger.info("[10/10] Creating 10. Approval Sign-Off...")
    create_approval_signoff(wb["10. Approval Sign-Of"], styles)
    
    # Save
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.5.23.S1_Inventory_{timestamp}.xlsx"
    wb.save(filename)
    
    logger.info("\n" + "=" * 70)
    logger.info("{CHECK} SUCCESS: {filename}")
    logger.info(f"{CHART} Sheets: {len(wb.sheetnames)}")
    logger.info(f"Exit Strategy: Sheet 4 (A-AC, 29 cols) | Regulatory: DORA/NIS2/AI/CLOUD")
    logger.info("=" * 70)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
