#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.23.S3 - Secure Configuration & Deployment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.23: Information Security for Use of Cloud Services
Assessment Domain 3 of 5: Secure Configuration & Deployment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific cloud architecture, security baselines, and
deployment requirements.

Key customization areas:
1. Security baseline standards (CIS benchmarks per your platforms)
2. Configuration requirements (IAM, encryption per your policies)
3. Network security controls (VPC, firewall rules per your architecture)
4. Deployment pipelines (CI/CD security gates per your SDLC)
5. Monitoring requirements (aligned with your SOC capabilities)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for cloud services)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

**Purpose:**
Enables systematic assessment of cloud security configurations and deployment
practices supporting ISO 27001:2022 Control A.5.23 requirements.

**Assessment Scope:**
- Security baseline compliance (CIS, cloud provider benchmarks)
- IAM configuration (least privilege, MFA, service accounts)
- Network security (VPC, security groups, network ACLs)
- Data protection (encryption at rest and in transit)
- Logging and monitoring configuration
- Deployment security (IaC security, CI/CD controls)
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance
2. Security Baseline - Benchmark compliance
3. IAM Configuration - Identity and access controls
4. Network Security - Perimeter and segmentation
5. Data Protection - Encryption and key management
6. Logging & Monitoring - Detection controls
7. Deployment Security - CI/CD and IaC
8. Gap Analysis - Configuration deficiencies
9. Evidence Register - Audit evidence tracking
10. Approval & Sign-Off - Stakeholder approval

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_reg_a523_3_secure_config.py

Requirements:
    - Python 3.8+
    - openpyxl library: pip install openpyxl

Output:
    ISMS-IMP-A.5.23.S3_SecureConfig_YYYYMMDD.xlsx
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.23.S3"
WORKBOOK_NAME = "Secure Configuration & Deployment"
CONTROL_ID = "A.5.19-23"
CONTROL_NAME = "Information Security for Use of Cloud Services"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
CLOUD = '\u2601'      # ☁️  Cloud
GLOBE = '\U0001F310'  # 🌐 Globe
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    sheets = [
        "Instructions & Legend",
        "2. Configuration Baseline",
        "3. Access Control Setup",
        "4. Network Security",
        "5. Encryption Configuration",
        "6. Deployment Checklist",
        "7. Jurisdictional Risk",
        "8. Summary Dashboard",         # Renumbered from 7
        "9. Evidence Register",         # Renumbered from 8
        "10. Approval Sign-Of",        # Renumbered from 9
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    
    return wb


def setup_styles() -> dict:
    """Define all cell styles - creates NEW objects to avoid shared object issues."""
    styles = {
        "header": {
            "font_params": {"name": "Calibri", "size": 14, "bold": True, "color": "FFFFFF"},
            "fill_params": {"start_color": "003366", "end_color": "003366", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "subheader": {
            "font_params": {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"},
            "fill_params": {"start_color": "4472C4", "end_color": "4472C4", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "column_header": {
            "font_params": {"name": "Calibri", "size": 10, "bold": True},
            "fill_params": {"start_color": "D9D9D9", "end_color": "D9D9D9", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "input_cell": {
            "fill_params": {"start_color": "FFFFCC", "end_color": "FFFFCC", "fill_type": "solid"},
            "alignment_params": {"horizontal": "left", "vertical": "center", "wrap_text": True},
        },
        "warning": {
            "font_params": {"name": "Calibri", "size": 10, "bold": True, "color": "FF0000"},
            "fill_params": {"start_color": "FFEB9C", "end_color": "FFEB9C", "fill_type": "solid"},
            "alignment_params": {"horizontal": "left", "vertical": "center", "wrap_text": True},
        },
    }
    return styles


def apply_style(cell, style_def, style_type):
    """Apply style to cell using NEW objects."""
    if "font_params" in style_def:
        cell.font = Font(**style_def["font_params"])
    if "fill_params" in style_def:
        cell.fill = PatternFill(**style_def["fill_params"])
    if "alignment_params" in style_def:
        cell.alignment = Alignment(**style_def["alignment_params"])
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: COLUMN DEFINITIONS (BASE + EXTENDED)
# ============================================================================

def get_base_config_columns() -> dict:
    """Return base 17 columns (A-Q) standard for ALL config assessment sheets."""
    return {
        "Assessment_ID": 14,
        "Cloud_Service_Name": 28,
        "Vendor_Name": 22,
        "Service_Type": 20,
        "Environment": 18,
        "Service_Criticality": 18,
        "Data_Classification": 20,
        "Configuration_Item": 30,
        "Status": 15,
        "Evidence_Location": 30,
        "Gap_Description": 35,
        "Remediation_Needed": 16,
        "Exception_ID": 14,
        "Risk_ID": 14,
        "Compensating_Controls": 30,
        "Responsible_Team": 20,
        "Target_Remediation_Date": 18,
    }


def get_extended_columns_config_baseline() -> dict:
    """Extended columns R-AA for Configuration Baseline (includes DORA/NIS2/AI Act)."""
    return {
        "Baseline_Version": 16,
        "Change_Mgmt_ID": 16,
        "Drift_Monitoring": 16,
        "CIS_Benchmark": 16,
        "Last_Validated": 16,
        "Config_Backup": 16,
        "Privileged_Access_Log": 18,
        "Security_By_Default": 16,
        # Regulatory columns (Y-AA)
        "DORA_Config_Documentation": 22,
        "NIS2_Secure_Deployment": 20,
        "AI_System_Deployment_Controls": 24,
    }


def get_extended_columns_access_control() -> dict:
    """Extended columns R-X for Access Control Setup."""
    return {
        "SSO_Integrated": 16,
        "MFA_Enforced": 16,
        "RBAC_Implemented": 16,
        "Privileged_Access_JIT": 18,
        "Service_Accounts_Inventoried": 20,
        "Last_Access_Review": 18,
        "Admin_Account_Count": 16,
    }


def get_extended_columns_network() -> dict:
    """Extended columns R-X for Network Security."""
    return {
        "IP_Restrictions": 16,
        "Private_Connectivity": 18,
        "Network_Segmentation": 18,
        "WAF_Enabled": 14,
        "DDoS_Protection": 16,
        "Firewall_Rules_Documented": 20,
        "Public_Endpoints_Count": 18,
    }


def get_extended_columns_encryption() -> dict:
    """Extended columns R-X for Encryption Configuration."""
    return {
        "Encryption_At_Rest": 18,
        "Encryption_In_Transit": 18,
        "Encryption_Algorithm": 18,
        "Key_Management": 20,
        "Key_Rotation_Period": 18,
        "HSM_Used": 14,
        "Last_Key_Rotation": 18,
    }


def get_extended_columns_deployment() -> dict:
    """Extended columns R-X for Deployment Checklist."""
    return {
        "Pre_Deploy_Scan": 16,
        "Security_Approved": 16,
        "Backup_Tested": 16,
        "Monitoring_Enabled": 18,
        "Runbook_Available": 16,
        "Rollback_Plan": 16,
        "Deployment_Date": 16,
    }


def get_jurisdictional_columns() -> dict:
    """Columns A-T for Jurisdictional Risk Assessment."""
    return {
        "Assessment_ID": 14,
        "Cloud_Service_Name": 25,
        "Provider_Name": 22,
        "Provider_HQ_Country": 18,
        "Provider_HQ_Jurisdiction": 20,
        "US_Parent_Company": 14,
        "CLOUD_Act_Applicability": 20,
        "Data_Processing_Locations": 25,
        "EU_Data_Boundary_Available": 18,
        "Customer_Managed_Keys": 16,
        "Legal_Challenge_Commitment": 18,
        "Adequacy_Decision_Status": 18,
        "Transfer_Mechanism": 16,
        "Risk_Level": 14,
        "Risk_Accepted_By": 18,
        "Risk_Acceptance_Date": 16,
        "Compensating_Controls": 28,
        "Review_Date": 14,
        "Evidence_Reference": 20,
        "Notes": 30,
    }


def get_all_columns(sheet_type: str) -> dict:
    """Combine base + extended columns based on sheet type."""
    base = get_base_config_columns()
    extended_map = {
        "config_baseline": get_extended_columns_config_baseline,
        "access_control": get_extended_columns_access_control,
        "network": get_extended_columns_network,
        "encryption": get_extended_columns_encryption,
        "deployment": get_extended_columns_deployment,
        "jurisdictional": get_jurisdictional_columns,
    }
    extended_func = extended_map.get(sheet_type)
    
    if sheet_type == "jurisdictional":
        # Jurisdictional sheet has its own column structure
        return extended_func()
    elif extended_func:
        return {**base, **extended_func()}
    return base


logger.info("Part 1 loaded: Foundation + Column Definitions")

#!/usr/bin/env python3
"""
PART 2: VALIDATION DEFINITIONS + CHECKLISTS
Depends on Part 1
"""

# ============================================================================
# SECTION 3: VALIDATION DEFINITIONS
# ============================================================================

def create_base_validations(ws) -> dict:
    """Create data validation objects for base columns."""
    validations = {}
    
    validations['service_type'] = DataValidation(
        type="list",
        formula1='"SaaS,IaaS,PaaS,Security Service,Cloud Storage,Other"',
        allow_blank=False
    )
    ws.add_data_validation(validations['service_type'])
    
    validations['environment'] = DataValidation(
        type="list",
        formula1='"Production,Staging,Development,Test,All"',
        allow_blank=False
    )
    ws.add_data_validation(validations['environment'])
    
    validations['criticality'] = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(validations['criticality'])
    
    validations['data_class'] = DataValidation(
        type="list",
        formula1='"Restricted,Confidential,Internal,Public,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['data_class'])
    
    validations['status'] = DataValidation(
        type="list",
        formula1=f'"{CHECK} Compliant,⚠️ Partial,❌ Non-Compliant,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['status'])
    
    validations['yes_no'] = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(validations['yes_no'])
    
    validations['responsible_team'] = DataValidation(
        type="list",
        formula1='"DevOpsSec,DevOps,Cloud Teams,IT Operations,Security"',
        allow_blank=False
    )
    ws.add_data_validation(validations['responsible_team'])
    
    return validations


def create_regulatory_validations(ws) -> dict:
    """Create regulatory-specific validations."""
    validations = {}
    
    validations['provider_hq_jurisdiction'] = DataValidation(
        type="list",
        formula1='"Switzerland,EU/EEA,United Kingdom,United States,Other Adequate Country,Non-Adequate Country"',
        allow_blank=False
    )
    ws.add_data_validation(validations['provider_hq_jurisdiction'])
    
    validations['cloud_act_exposure'] = DataValidation(
        type="list",
        formula1='"No Exposure,Potential Exposure (US HQ),Mitigated (EU Data Boundary),Mitigated (Encryption + Key Control),Accepted Risk (Documented),Under Assessment"',
        allow_blank=False
    )
    ws.add_data_validation(validations['cloud_act_exposure'])
    
    validations['transfer_mechanism'] = DataValidation(
        type="list",
        formula1='"SCCs,BCRs,Adequacy Decision,None,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['transfer_mechanism'])
    
    validations['risk_level'] = DataValidation(
        type="list",
        formula1='"Low,Medium,High,Critical"',
        allow_blank=False
    )
    ws.add_data_validation(validations['risk_level'])
    
    validations['yes_no_partial_unknown'] = DataValidation(
        type="list",
        formula1='"Yes,No,Partial,Unknown"',
        allow_blank=False
    )
    ws.add_data_validation(validations['yes_no_partial_unknown'])
    
    validations['yes_no_planned'] = DataValidation(
        type="list",
        formula1='"Yes,No,Planned"',
        allow_blank=False
    )
    ws.add_data_validation(validations['yes_no_planned'])
    
    validations['dora_compliance'] = DataValidation(
        type="list",
        formula1='"Full Compliance,Partial Compliance,Non-Compliant,N/A (Not in scope)"',
        allow_blank=False
    )
    ws.add_data_validation(validations['dora_compliance'])
    
    validations['nis2_deployment'] = DataValidation(
        type="list",
        formula1='"Compliant,Partial (In Progress),Non-Compliant,N/A (Not in scope)"',
        allow_blank=False
    )
    ws.add_data_validation(validations['nis2_deployment'])
    
    validations['ai_system'] = DataValidation(
        type="list",
        formula1='"No AI Systems,Low-Risk AI Only,High-Risk AI (Assessed),High-Risk AI (Assessment Pending),N/A"',
        allow_blank=False
    )
    ws.add_data_validation(validations['ai_system'])
    
    return validations


# ============================================================================
# SECTION 4: CHECKLIST DEFINITIONS
# ============================================================================

def get_checklist_config_baseline() -> list:
    """Configuration Baseline checklist (INCLUDES DORA/NIS2/AI Act items)."""
    return [
        # Original items
        ("Configuration baseline documented and version-controlled", "Baseline doc + version history"),
        ("Change management process for cloud configurations", "Change log/ITSM tickets"),
        ("Configuration drift monitoring enabled", "Drift detection config"),
        ("Security configuration validated against CIS benchmarks", "CIS scan report"),
        ("Configuration backup and recovery tested", "Backup test results"),
        ("Privileged access to configurations logged and monitored", "Access log screenshot"),
        ("Security-by-default configurations applied", "Default settings review"),
        
        # DORA items
        ("CFG-DORA-01: Configuration baseline documented and version-controlled", "DORA: Version control"),
        ("CFG-DORA-02: Change management process for cloud configurations", "DORA: Change records"),
        ("CFG-DORA-03: Configuration drift monitoring enabled", "DORA: Monitoring config"),
        ("CFG-DORA-04: Security configuration validated against CIS benchmarks", "DORA: CIS compliance"),
        ("CFG-DORA-05: Configuration backup and recovery tested", "DORA: Test results"),
        ("CFG-DORA-06: Privileged access to configurations logged and monitored", "DORA: Access logs"),
        
        # NIS2 items
        ("CFG-NIS2-01: Security-by-default configurations applied", "NIS2: Default settings"),
        ("CFG-NIS2-02: Network segmentation implemented", "NIS2: Network diagram"),
        ("CFG-NIS2-03: Multi-factor authentication enforced for admin access", "NIS2: MFA config"),
        ("CFG-NIS2-04: Vulnerability scanning configured pre-deployment", "NIS2: Scan config"),
        
        # AI Act items
        ("CFG-AI-01: AI system risk classification documented", "AI Act: Risk assessment"),
        ("CFG-AI-02: High-risk AI systems have conformity assessment", "AI Act: Conformity cert"),
        ("CFG-AI-03: AI system transparency notices configured", "AI Act: Notice config"),
        ("CFG-AI-04: AI system logging and monitoring enabled", "AI Act: AI logs"),
        ("CFG-AI-05: Human oversight mechanisms configured", "AI Act: Oversight config"),
    ]


def get_checklist_access_control() -> list:
    """Access Control Setup checklist."""
    return [
        ("SSO configured via organisational IdP (Entra ID, Okta, etc.)", "SSO config screenshot"),
        ("MFA enforced for all users", "MFA policy screenshot"),
        ("MFA mandatory for privileged accounts", "Admin MFA verification"),
        ("RBAC roles documented and implemented", "Role matrix document"),
        ("Least privilege principle applied", "Permission audit"),
        ("Service accounts inventoried with owners", "Service account register"),
        ("Privileged access is time-limited (JIT)", "PAM configuration"),
        ("Access reviews performed quarterly", "Access review report"),
        ("Default accounts disabled or renamed", "Security baseline check"),
        ("Failed login alerting configured", "Alert rule screenshot"),
    ]


def get_checklist_network() -> list:
    """Network Security checklist."""
    return [
        ("IP allowlisting configured for admin access", "Firewall rules"),
        ("Private endpoints used where available", "Network architecture"),
        ("VPN/ExpressRoute for sensitive workloads", "Connectivity config"),
        ("Network segmentation between environments", "Network diagram"),
        ("WAF configured for web applications", "WAF rules export"),
        ("DDoS protection enabled", "DDoS settings"),
        ("Firewall rules documented and reviewed", "Rule documentation"),
        ("Public endpoints minimized and justified", "Endpoint inventory"),
    ]


def get_checklist_encryption() -> list:
    """Encryption Configuration checklist."""
    return [
        ("TLS 1.2+ enforced for all connections", "TLS config, SSL scan"),
        ("AES-256 encryption at rest enabled", "Encryption settings"),
        ("Encryption keys stored securely (HSM/KMS)", "Key management config"),
        ("Key rotation policy implemented", "Rotation schedule"),
        ("Customer-managed keys (CMK) used for sensitive data", "CMK configuration"),
        ("Key access logged and monitored", "Key access logs"),
        ("Encryption validated for backups", "Backup encryption check"),
        ("Secure deletion process documented", "Deletion procedures"),
    ]


def get_checklist_deployment() -> list:
    """Deployment Checklist items."""
    return [
        ("Pre-deployment security scan completed", "Scan report"),
        ("Security team approval obtained", "Approval ticket"),
        ("Backup and recovery tested", "Test results"),
        ("Monitoring and alerting configured", "Monitoring config"),
        ("Deployment runbook available", "Runbook document"),
        ("Rollback plan documented and tested", "Rollback procedure"),
        ("Production deployment authorized", "Change approval"),
        ("Post-deployment validation completed", "Validation report"),
    ]


def get_checklist_jurisdictional() -> list:
    """Jurisdictional Risk Assessment checklist."""
    return [
        ("Provider HQ jurisdiction identified and documented", "Vendor documentation"),
        ("US parent company status verified", "Corporate structure review"),
        ("CLOUD Act exposure assessed for US-nexus providers", "Legal review"),
        ("Data processing locations documented in DPA", "DPA review"),
        ("EU Data Boundary availability checked", "Vendor announcement/docs"),
        ("Customer-managed encryption keys option evaluated", "Technical documentation"),
        ("Legal challenge commitment verified", "Contract/Public statement"),
        ("Transfer mechanism documented (SCCs, BCRs, etc.)", "DPA"),
        ("Risk acceptance recorded if residual risk remains", "Risk register"),
        ("Compensating controls documented for high-risk providers", "Control documentation"),
    ]


logger.info("Part 2 loaded: Validations + Checklists (Regulatory Enhanced)")

# ============================================================================
# SECTION 5: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet with regulatory compliance info."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header — single merged row with two-line title
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.5.23.S3  -  Secure Configuration & Deployment\n"
        "ISO/IEC 27001:2022 - Control A.5.23: Information Security for Use of Cloud Services"
    )
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 40

    # Document Information
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.5.23.S3"),
        ("Assessment Area", "Secure Configuration & Deployment"),
        ("Related Policy", "ISMS-POL-A.5.19-23-S3, S5"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        row += 1

    # Regulatory Compliance Section
    row += 1
    ws[f"A{row}"] = "REGULATORY COMPLIANCE"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="003366")

    row += 1
    reg_info = [
        "This workbook includes regulatory compliance tracking for:",
        f"{BULLET} DORA (Digital Operational Resilience Act) - Configuration documentation",
        f"{BULLET} NIS2 (Network and Information Security Directive 2) - Secure deployment",
        f"{BULLET} EU AI Act - AI system deployment controls",
        f"{BULLET} US CLOUD Act - Jurisdictional risk for US-headquartered providers",
        "",
        "Regulatory columns are included in the Configuration Baseline sheet (Y-AA).",
        "Complete only the regulatory sections applicable to your organisation.",
    ]
    for line in reg_info:
        ws[f"A{row}"] = line
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        row += 1

    # How to Use This Workbook
    row += 1
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Reference ISMS-IMP-A.5.23.S1 (Inventory) for authoritative cloud service list.",
        "2. Complete each worksheet tab (2-7) for all cloud services requiring config assessment.",
        "3. Use dropdown menus for standardised entries (Environment, Status, etc.).",
        "4. Fill in yellow-highlighted cells with your configuration-specific information.",
        "5. Complete Jurisdictional Risk sheet (7) for all US-nexus providers.",
        "6. Document DORA configuration requirements if in scope.",
        "7. Validate NIS2 secure deployment procedures if applicable.",
        "8. Complete AI Act deployment controls for AI-enabled services.",
        "9. Provide evidence: screenshots, config exports, security scan results.",
        "10. Summary Dashboard auto-calculates compliance by config area and environment.",
        "11. Obtain final approval from IT Operations, Security, DPO, and CISO.",
    ]
    row += 1
    for instr in instructions:
        ws[f"A{row}"] = instr
        row += 1

    # Status Legend — proper table with headers
    row += 1
    ws[f"A{row}"] = "STATUS LEGEND"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    for ci, hdr in enumerate(("Symbol", "Status", "Description"), start=1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    legend = [
        (CHECK, "Compliant", "Configuration meets security requirements", "C6EFCE"),
        (WARNING, "Partial", "Some settings correct, gaps exist", "FFEB9C"),
        (XMARK, "Non-Compliant", "Configuration does not meet requirements", "FFC7CE"),
        ("N/A", "Not Applicable", "Control does not apply to this service", None),
    ]

    row += 1
    for sym, status, desc, color in legend:
        ws.cell(row=row, column=1, value=sym).border = border
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if color:
            s.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

    # Acceptable Evidence
    row += 1
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    evidence_types = [
        "✓ Admin console screenshots (with timestamps)",
        "✓ Configuration export files (JSON, YAML, XML)",
        "✓ Security posture reports (Azure Secure Score, AWS Security Hub)",
        "✓ Configuration baseline documentation (version-controlled)",
        "✓ CIS benchmark compliance report",
        "✓ Configuration drift monitoring screenshots",
        "✓ Network segmentation diagrams",
        "✓ AI system risk classification (if applicable)",
        "✓ AI conformity assessment (for high-risk systems)",
        "✓ Jurisdictional risk assessment (for US-nexus providers)",
        "✓ Risk acceptance form (signed by CISO/DPO)",
        "✓ Data Processing Agreement (DPA) with SCCs/BCRs",
    ]
    row += 1
    for ev in evidence_types:
        ws[f"A{row}"] = ev
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 6: GENERIC ASSESSMENT SHEET ENGINE
# ============================================================================

def create_config_assessment_sheet(ws, styles, section_title, policy_req,
                                   question, sheet_type, checklist_items):
    """Generic assessment sheet creator for configuration controls."""
    
    # Determine total columns based on sheet type
    all_columns = get_all_columns(sheet_type)
    total_cols = len(all_columns)
    last_col_letter = get_column_letter(total_cols)
    
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"{section_title}\n{policy_req}"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 50
    
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"Assessment Question: {question}"
    apply_style(ws["A2"], styles["subheader"], "subheader")
    ws.row_dimensions[2].height = 30
    
    col_names = list(all_columns.keys())
    col_widths = list(all_columns.values())
    
    row = 4
    for col_idx, (col_name, col_width) in enumerate(zip(col_names, col_widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        apply_style(cell, styles["column_header"], "header")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # Data rows 5-30 (26 entries)
    for data_row in range(5, 31):
        # Auto-generate Assessment_ID in column A
        ws.cell(row=data_row, column=1, value=f'=CONCATENATE("{sheet_type[:3].upper()}-",TEXT(ROW()-4,"000"))')
        
        for col_idx in range(2, len(col_names) + 1):
            cell = ws.cell(row=data_row, column=col_idx, value="")
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Apply validations
    if sheet_type != "jurisdictional":
        base_vals = create_base_validations(ws)
        base_vals['service_type'].add("D5:D30")
        base_vals['environment'].add("E5:E30")
        base_vals['criticality'].add("F5:F30")
        base_vals['data_class'].add("G5:G30")
        base_vals['status'].add("I5:I30")
        base_vals['yes_no'].add("L5:L30")
        base_vals['responsible_team'].add("P5:P30")
    
    _apply_extended_validations(ws, sheet_type)
    
    # Checklist section
    checklist_row = 33
    ws[f"A{checklist_row}"] = "COMPLIANCE CHECKLIST"
    ws[f"A{checklist_row}"].font = Font(bold=True, size=11)
    
    checklist_row += 1
    ws[f"A{checklist_row}"] = "☐"
    ws[f"B{checklist_row}"] = "Requirement"
    ws[f"C{checklist_row}"] = "Evidence"
    for col in ["A", "B", "C"]:
        ws[f"{col}{checklist_row}"].font = Font(bold=True)
    
    checklist_row += 1
    for req, evidence in checklist_items:
        ws[f"A{checklist_row}"] = "☐"
        ws[f"B{checklist_row}"] = req
        ws[f"C{checklist_row}"] = evidence
        checklist_row += 1
    
    ws.freeze_panes = "A5"


def _apply_extended_validations(ws, sheet_type):
    """Apply extended column validations based on sheet type."""
    
    if sheet_type == "config_baseline":
        _create_config_baseline_validations(ws)
    elif sheet_type == "access_control":
        _create_access_control_validations(ws)
    elif sheet_type == "network":
        _create_network_validations(ws)
    elif sheet_type == "encryption":
        _create_encryption_validations(ws)
    elif sheet_type == "deployment":
        _create_deployment_validations(ws)
    elif sheet_type == "jurisdictional":
        _create_jurisdictional_validations(ws)


def _create_config_baseline_validations(ws):
    """Configuration Baseline extended validations (includes DORA/NIS2/AI Act)."""
    dv_yes_no = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=False)
    ws.add_data_validation(dv_yes_no)
    dv_yes_no.add("R5:R30")  # Baseline_Version
    dv_yes_no.add("T5:T30")  # Drift_Monitoring
    dv_yes_no.add("U5:U30")  # CIS_Benchmark
    dv_yes_no.add("V5:V30")  # Config_Backup
    dv_yes_no.add("W5:W30")  # Privileged_Access_Log
    dv_yes_no.add("X5:X30")  # Security_By_Default
    
    # DORA compliance (column Y)
    reg_vals = create_regulatory_validations(ws)
    reg_vals['dora_compliance'].add("Y5:Y30")
    
    # NIS2 deployment (column Z)
    reg_vals['nis2_deployment'].add("Z5:Z30")
    
    # AI System deployment (column AA)
    reg_vals['ai_system'].add("AA5:AA30")


def _create_access_control_validations(ws):
    """Access Control Setup extended validations."""
    dv_sso = DataValidation(type="list", formula1='"Yes,No,Partial,N/A"', allow_blank=False)
    ws.add_data_validation(dv_sso)
    dv_sso.add("R5:R30")
    
    dv_mfa = DataValidation(type="list", formula1='"Yes (All Users),Yes (Admins Only),No,N/A"', allow_blank=False)
    ws.add_data_validation(dv_mfa)
    dv_mfa.add("S5:S30")
    
    dv_rbac = DataValidation(type="list", formula1='"Yes,Partial,No,N/A"', allow_blank=False)
    ws.add_data_validation(dv_rbac)
    dv_rbac.add("T5:T30")
    
    dv_jit = DataValidation(type="list", formula1='"Yes,No,Planned,N/A"', allow_blank=False)
    ws.add_data_validation(dv_jit)
    dv_jit.add("U5:U30")
    
    dv_svc = DataValidation(type="list", formula1='"Yes,Partial,No,Unknown"', allow_blank=False)
    ws.add_data_validation(dv_svc)
    dv_svc.add("V5:V30")


def _create_network_validations(ws):
    """Network Security extended validations."""
    dv_ip = DataValidation(type="list", formula1='"Yes (Allowlist),Yes (Geo),No,N/A"', allow_blank=False)
    ws.add_data_validation(dv_ip)
    dv_ip.add("R5:R30")
    
    dv_private = DataValidation(type="list", formula1='"Yes (Private Link),Yes (VPN),Public Only,N/A"', allow_blank=False)
    ws.add_data_validation(dv_private)
    dv_private.add("S5:S30")
    
    dv_seg = DataValidation(type="list", formula1='"Yes,Partial,No,N/A"', allow_blank=False)
    ws.add_data_validation(dv_seg)
    dv_seg.add("T5:T30")
    
    dv_waf = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=False)
    ws.add_data_validation(dv_waf)
    dv_waf.add("U5:U30")
    
    dv_ddos = DataValidation(type="list", formula1='"Yes (Advanced),Yes (Basic),No,N/A"', allow_blank=False)
    ws.add_data_validation(dv_ddos)
    dv_ddos.add("V5:V30")
    
    dv_fw = DataValidation(type="list", formula1='"Yes,Partial,No,N/A"', allow_blank=False)
    ws.add_data_validation(dv_fw)
    dv_fw.add("W5:W30")


def _create_encryption_validations(ws):
    """Encryption Configuration extended validations."""
    dv_enc_rest = DataValidation(type="list", formula1='"Yes (Provider Key),Yes (CMK),No,N/A"', allow_blank=False)
    ws.add_data_validation(dv_enc_rest)
    dv_enc_rest.add("R5:R30")
    
    dv_enc_transit = DataValidation(type="list", formula1='"Yes (TLS 1.3),Yes (TLS 1.2),No,N/A"', allow_blank=False)
    ws.add_data_validation(dv_enc_transit)
    dv_enc_transit.add("S5:S30")
    
    dv_algo = DataValidation(type="list", formula1='"AES-256,AES-128,ChaCha20,Other,Unknown"', allow_blank=False)
    ws.add_data_validation(dv_algo)
    dv_algo.add("T5:T30")
    
    dv_key = DataValidation(type="list", formula1='"Provider Managed,Customer Managed (HSM),Customer Managed (Software),N/A"', allow_blank=False)
    ws.add_data_validation(dv_key)
    dv_key.add("U5:U30")
    
    dv_rotation = DataValidation(type="list", formula1='"90 days,180 days,365 days,Manual,N/A"', allow_blank=False)
    ws.add_data_validation(dv_rotation)
    dv_rotation.add("V5:V30")
    
    dv_hsm = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=False)
    ws.add_data_validation(dv_hsm)
    dv_hsm.add("W5:W30")


def _create_deployment_validations(ws):
    """Deployment Checklist extended validations."""
    dv_yes_no = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=False)
    ws.add_data_validation(dv_yes_no)
    dv_yes_no.add("R5:R30")  # Pre_Deploy_Scan
    dv_yes_no.add("S5:S30")  # Security_Approved
    dv_yes_no.add("T5:T30")  # Backup_Tested
    dv_yes_no.add("U5:U30")  # Monitoring_Enabled
    dv_yes_no.add("V5:V30")  # Runbook_Available
    dv_yes_no.add("W5:W30")  # Rollback_Plan


def _create_jurisdictional_validations(ws):
    """Jurisdictional Risk Assessment validations."""
    reg_vals = create_regulatory_validations(ws)
    
    # Column E - Provider HQ Jurisdiction
    reg_vals['provider_hq_jurisdiction'].add("E5:E30")
    
    # Column F - US Parent Company
    dv_yes_no_unknown = DataValidation(type="list", formula1='"Yes,No,Unknown"', allow_blank=False)
    ws.add_data_validation(dv_yes_no_unknown)
    dv_yes_no_unknown.add("F5:F30")
    
    # Column G - CLOUD Act Applicability
    reg_vals['cloud_act_exposure'].add("G5:G30")
    
    # Column I - EU Data Boundary Available
    reg_vals['yes_no_planned'].add("I5:I30")
    
    # Column J - Customer Managed Keys
    reg_vals['yes_no_planned'].add("J5:J30")
    
    # Column K - Legal Challenge Commitment
    reg_vals['yes_no_partial_unknown'].add("K5:K30")
    
    # Column M - Transfer Mechanism
    reg_vals['transfer_mechanism'].add("M5:M30")
    
    # Column N - Risk Level
    reg_vals['risk_level'].add("N5:N30")


# ============================================================================
# SECTION 7: INDIVIDUAL SHEET CREATORS
# ============================================================================

def create_2_configuration_baseline(ws, styles):
    """Sheet 2: Configuration Baseline (with DORA/NIS2/AI Act columns Y-AA)."""
    create_config_assessment_sheet(
        ws=ws, styles=styles,
        section_title="CONFIGURATION BASELINE",
        policy_req="Policy Requirement: Configuration baseline documented, change-controlled, drift-monitored, CIS-validated (Policy S5 Section 9.1)",
        question="Are configuration baselines properly documented, version-controlled, and validated for your cloud services?",
        sheet_type="config_baseline",
        checklist_items=get_checklist_config_baseline()
    )


def create_3_access_control(ws, styles):
    """Sheet 3: Access Control Setup."""
    create_config_assessment_sheet(
        ws=ws, styles=styles,
        section_title="ACCESS CONTROL SETUP",
        policy_req="Policy Requirement: SSO integration, MFA enforced, RBAC implemented, privileged access controlled (Policy S5 Section 9.1)",
        question="Are identity and access controls properly configured for your cloud services?",
        sheet_type="access_control",
        checklist_items=get_checklist_access_control()
    )


def create_4_network_security(ws, styles):
    """Sheet 4: Network Security."""
    create_config_assessment_sheet(
        ws=ws, styles=styles,
        section_title="NETWORK SECURITY CONFIGURATION",
        policy_req="Policy Requirement: Access restrictions, secure connectivity, network segmentation (Policy S5 Section 5.2)",
        question="Are network security controls properly configured for your cloud services?",
        sheet_type="network",
        checklist_items=get_checklist_network()
    )


def create_5_encryption_config(ws, styles):
    """Sheet 5: Encryption Configuration."""
    create_config_assessment_sheet(
        ws=ws, styles=styles,
        section_title="ENCRYPTION CONFIGURATION",
        policy_req="Policy Requirement: Encryption at rest and in transit, key management, secure key storage (Policy S5 Section 9.2)",
        question="Are encryption controls properly configured for your cloud services?",
        sheet_type="encryption",
        checklist_items=get_checklist_encryption()
    )


def create_6_deployment_checklist(ws, styles):
    """Sheet 6: Deployment Checklist."""
    create_config_assessment_sheet(
        ws=ws, styles=styles,
        section_title="DEPLOYMENT CHECKLIST",
        policy_req="Policy Requirement: Pre-deployment validation, security approval, backup testing, monitoring (Policy S5 Section 5.2)",
        question="Have all deployment security requirements been met for your cloud services?",
        sheet_type="deployment",
        checklist_items=get_checklist_deployment()
    )


def create_7_jurisdictional_risk(ws, styles):
    """Sheet 7 - Jurisdictional Risk Assessment (CLOUD Act)."""
    
    # Title with warning
    ws.merge_cells("A1:T1")
    ws["A1"] = "JURISDICTIONAL RISK ASSESSMENT"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells("A2:T2")
    ws["A2"] = "CLOUD Act, Data Sovereignty, and Cross-Border Transfer Analysis"
    apply_style(ws["A2"], styles["subheader"], "subheader")
    ws.row_dimensions[2].height = 25
    
    # Warning row
    ws.merge_cells("A3:T3")
    ws["A3"] = f"{WARNING} US-headquartered providers may be compelled to disclose data under CLOUD Act regardless of data location."
    apply_style(ws["A3"], styles["warning"], "warning")
    ws.row_dimensions[3].height = 30
    
    # Column headers
    all_columns = get_jurisdictional_columns()
    col_names = list(all_columns.keys())
    col_widths = list(all_columns.values())
    
    row = 5
    for col_idx, (col_name, col_width) in enumerate(zip(col_names, col_widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        apply_style(cell, styles["column_header"], "header")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # Data rows 6-30 (25 entries)
    for data_row in range(6, 31):
        # Auto-generate Assessment_ID in column A
        ws.cell(row=data_row, column=1, value=f'=CONCATENATE("JRA-",TEXT(ROW()-5,"000"))')
        
        for col_idx in range(2, len(col_names) + 1):
            cell = ws.cell(row=data_row, column=col_idx, value="")
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Apply validations
    _create_jurisdictional_validations(ws)
    
    # Checklist
    checklist_row = 33
    ws[f"A{checklist_row}"] = "JURISDICTIONAL RISK CHECKLIST"
    ws[f"A{checklist_row}"].font = Font(bold=True, size=11)
    
    checklist_row += 1
    ws[f"A{checklist_row}"] = "☐"
    ws[f"B{checklist_row}"] = "Requirement"
    ws[f"C{checklist_row}"] = "Evidence"
    for col in ["A", "B", "C"]:
        ws[f"{col}{checklist_row}"].font = Font(bold=True)
    
    checklist_row += 1
    for req, evidence in get_checklist_jurisdictional():
        ws[f"A{checklist_row}"] = "☐"
        ws[f"B{checklist_row}"] = req
        ws[f"C{checklist_row}"] = evidence
        checklist_row += 1
    
    ws.freeze_panes = "A6"


logger.info("Part 3 loaded: Sheet Creators (Instructions + Assessment Sheets 2-7)")

#!/usr/bin/env python3
"""
PART 4: DASHBOARD, EVIDENCE REGISTER, APPROVAL & MAIN EXECUTION
Depends on Parts 1, 2 & 3
"""

# ============================================================================
# SECTION 8: SUMMARY DASHBOARD
# ============================================================================

def create_8_summary_dashboard(ws, styles):
    """Sheet 8: Summary Dashboard with jurisdictional and regulatory compliance tables."""
    
    ws.merge_cells("A1:G1")
    ws["A1"] = "SECURE CONFIGURATION - COMPLIANCE SUMMARY DASHBOARD"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 35
    
    # TABLE 1: Compliance by Configuration Area
    row = 3
    ws[f"A{row}"] = "TABLE 1: COMPLIANCE BY CONFIGURATION AREA"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    
    row += 1
    headers = ["Configuration Area", "Total Items", f"{CHECK} Compliant", f"{WARNING} Partial", f"{XMARK} Non-Compliant", "N/A", "Compliance %"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        apply_style(cell, styles["column_header"], "header")
    
    areas = [
        ("Configuration Baseline", "'2. Configuration Baseline'"),
        ("Access Control Setup", "'3. Access Control Setup'"),
        ("Network Security", "'4. Network Security'"),
        ("Encryption Configuration", "'5. Encryption Configuration'"),
        ("Deployment Checklist", "'6. Deployment Checklist'"),
    ]
    
    row += 1
    start_data_row = row
    for area_name, sheet_ref in areas:
        ws.cell(row=row, column=1, value=area_name)
        ws.cell(row=row, column=2, value=f'=COUNTA({sheet_ref}!B5:B30)-COUNTBLANK({sheet_ref}!B5:B30)')
        ws.cell(row=row, column=3, value=f'=COUNTIF({sheet_ref}!I5:I30,"{CHECK}*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({sheet_ref}!I5:I30,"{WARNING}*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({sheet_ref}!I5:I30,"{XMARK}*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({sheet_ref}!I5:I30,"N/A")')
        ws.cell(row=row, column=7, value=f'=IF(B{row}-F{row}>0,ROUND(C{row}/(B{row}-F{row})*100,1)&"%","N/A")')
        row += 1
    
    end_data_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL")
    ws[f"A{row}"].font = Font(bold=True)
    for col in range(2, 7):
        ws.cell(row=row, column=col, value=f'=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{end_data_row})')
        ws.cell(row=row, column=col).font = Font(bold=True)
    ws.cell(row=row, column=7, value=f'=IF(B{row}-F{row}>0,ROUND(C{row}/(B{row}-F{row})*100,1)&"%","N/A")')
    ws.cell(row=row, column=7).font = Font(bold=True)
    
    # TABLE 2: Jurisdictional & CLOUD Act Risk Summary
    row += 3
    ws[f"A{row}"] = "TABLE 2: JURISDICTIONAL & CLOUD ACT RISK SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    
    row += 1
    metric_headers = ["Metric", "Count", "Status"]
    for col_idx, h in enumerate(metric_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        apply_style(cell, styles["column_header"], "header")
    
    jurisdictional_metrics = [
        ("US-HQ Providers (CLOUD Act Scope)", "=COUNTIF('7. Jurisdictional Risk'!E6:E30,\"United States\")", "=IF(B{row}>0,\"{WARNING} Review\",\"{CHECK} OK\")"),
        ("Providers with US Parent Company", "=COUNTIF('7. Jurisdictional Risk'!F6:F30,\"Yes\")", "=IF(B{row}>0,\"{WARNING} Review\",\"{CHECK} OK\")"),
        ("CLOUD Act Potential Exposure (Unmitigated)", "=COUNTIF('7. Jurisdictional Risk'!G6:G30,\"Potential*\")", "=IF(B{row}>0,\"{WARNING} Review\",\"{CHECK} OK\")"),
        ("CLOUD Act Mitigated", "=COUNTIFS('7. Jurisdictional Risk'!G6:G30,\"Mitigated*\")", "Info"),
        ("High/Critical Jurisdictional Risk", "=COUNTIF('7. Jurisdictional Risk'!N6:N30,\"High\")+COUNTIF('7. Jurisdictional Risk'!N6:N30,\"Critical\")", "=IF(B{row}>0,\"{XMARK} Action\",\"{CHECK} OK\")"),
        ("Providers Without EU Data Boundary", "=COUNTIF('7. Jurisdictional Risk'!I6:I30,\"No\")", "=IF(B{row}>0,\"{WARNING} Review\",\"{CHECK} OK\")"),
        ("Providers Without Customer-Managed Keys", "=COUNTIF('7. Jurisdictional Risk'!J6:J30,\"No\")", "=IF(B{row}>0,\"{WARNING} Review\",\"{CHECK} OK\")"),
    ]
    
    row += 1
    for metric_name, count_formula, status_formula in jurisdictional_metrics:
        ws.cell(row=row, column=1, value=metric_name)
        ws.cell(row=row, column=2, value=count_formula)
        
        # Replace {row} placeholder with actual row number
        status_formula_final = status_formula.replace("{row}", str(row))
        ws.cell(row=row, column=3, value=status_formula_final)
        row += 1
    
    # TABLE 3: Regulatory Deployment Compliance
    row += 2
    ws[f"A{row}"] = "TABLE 3: REGULATORY DEPLOYMENT COMPLIANCE"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    
    row += 1
    for col_idx, h in enumerate(metric_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        apply_style(cell, styles["column_header"], "header")
    
    regulatory_metrics = [
        ("DORA Non-Compliant Configurations", "=COUNTIF('2. Configuration Baseline'!Y5:Y30,\"Non-Compliant\")", "=IF(B{row}>0,\"{XMARK} Action\",\"{CHECK} OK\")"),
        ("NIS2 Non-Compliant Deployments", "=COUNTIF('2. Configuration Baseline'!Z5:Z30,\"Non-Compliant\")", "=IF(B{row}>0,\"{XMARK} Action\",\"{CHECK} OK\")"),
        ("High-Risk AI Systems Pending Assessment", "=COUNTIF('2. Configuration Baseline'!AA5:AA30,\"*Pending*\")", "=IF(B{row}>0,\"{WARNING} Review\",\"{CHECK} OK\")"),
    ]
    
    row += 1
    for metric_name, count_formula, status_formula in regulatory_metrics:
        ws.cell(row=row, column=1, value=metric_name)
        ws.cell(row=row, column=2, value=count_formula)
        
        status_formula_final = status_formula.replace("{row}", str(row))
        ws.cell(row=row, column=3, value=status_formula_final)
        row += 1
    
    # TABLE 4: Critical Gaps by Environment
    row += 3
    ws[f"A{row}"] = "TABLE 4: CRITICAL GAPS BY ENVIRONMENT"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    
    row += 1
    gap_headers = ["Gap Type", "Count", "Priority"]
    for col_idx, h in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        apply_style(cell, styles["column_header"], "header")
    
    gaps = [
        ("Production + Non-Compliant", "[Manual Count]", "Critical"),
        ("Critical Service + Non-Compliant", "[Manual Count]", "Critical"),
        ("MFA Not Enforced (Production)", "[Manual Count]", "High"),
        ("No Encryption at Rest", "[Manual Count]", "High"),
        ("Logging Not Enabled", "[Manual Count]", "High"),
        ("No Backup for Critical Services", "[Manual Count]", "High"),
    ]
    
    row += 1
    for gap, count, priority in gaps:
        ws.cell(row=row, column=1, value=gap)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=priority)
        
        if priority == "Critical":
            ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            ws.cell(row=row, column=3).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        row += 1
    
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 9: EVIDENCE REGISTER
# ============================================================================

def create_9_evidence_register(ws, styles):
    """Sheet 9: Evidence Register with regulatory evidence types."""

    ws.merge_cells("A1:I1")
    ws["A1"] = "EVIDENCE REGISTER - SECURE CONFIGURATION & DEPLOYMENT"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:I2")
    ws["A2"] = "Track all configuration evidence, scan reports, and compliance documentation"
    apply_style(ws["A2"], styles["subheader"], "subheader")

    columns = [
        ("Evidence ID", 18),
        ("Cloud Service Name", 25),
        ("Configuration Area", 22),
        ("Evidence Type", 25),
        ("Description", 35),
        ("File Location", 35),
        ("Capture Date", 16),
        ("Captured By", 18),
        ("Status", 14),
    ]

    row = 4
    for col_idx, (col_name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        apply_style(cell, styles["column_header"], "header")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows 5-104 (100 entries)
    for data_row in range(5, 105):
        i = data_row - 4
        ws.cell(row=data_row, column=1, value=f"EV-SCD-{i:03d}")
        for col_idx in range(2, 10):
            cell = ws.cell(row=data_row, column=col_idx, value="")
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Validation for Configuration Area
    dv_area = DataValidation(
        type="list",
        formula1='"Configuration Baseline,Access Control,Network Security,Encryption,Deployment,Jurisdictional Risk,DORA Compliance,NIS2 Compliance,AI Act Compliance"',
        allow_blank=False
    )
    ws.add_data_validation(dv_area)
    dv_area.add("C5:C104")

    # Validation for Evidence Type (expanded for regulatory)
    dv_type = DataValidation(
        type="list",
        formula1='"Screenshot,Config Export,Scan Report,Test Result,IaC Template,API Query,DPA with SCCs,Risk Assessment,Conformity Certificate,Legal Review,Vendor Documentation,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)
    dv_type.add("D5:D104")

    # Validation for Status
    dv_status = DataValidation(
        type="list",
        formula1='"Current,Outdated,Pending"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)
    dv_status.add("I5:I104")

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 10: APPROVAL SIGN-OFF
# ============================================================================

def create_10_approval_signoff(ws, styles):
    """Sheet 10: Multi-stakeholder Approval Sign-Off with DPO section."""
    
    ws.merge_cells("A1:E1")
    ws["A1"] = "APPROVAL SIGN-OFF - SECURE CONFIGURATION ASSESSMENT"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 35
    
    row = 3
    
    # ASSESSMENT SUMMARY (with regulatory metrics)
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")

    summary_fields = [
        ("Assessment Completed By:", "", "input"),
        ("Completion Date:", "", "input"),
        ("Total Configurations Assessed:", "='8. Summary Dashboard'!B9", "formula"),
        ("Overall Compliance %:", "='8. Summary Dashboard'!G9", "formula"),
        ("Jurisdictional Risks Identified:", "='8. Summary Dashboard'!B20", "formula"),
        ("DORA Config Non-Compliance:", "='8. Summary Dashboard'!B27", "formula"),
        ("NIS2 Deployment Non-Compliance:", "='8. Summary Dashboard'!B28", "formula"),
        ("Critical Gaps Identified:", "", "input"),
        ("Remediation Plan Attached:", "", "dropdown"),
    ]
    
    row += 1
    remediation_plan_row = None
    for label, value, field_type in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        
        if field_type == "formula":
            ws[f"B{row}"] = value
        elif field_type == "input":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        elif field_type == "dropdown":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            remediation_plan_row = row
        row += 1
    
    # Add dropdown for Remediation Plan
    if remediation_plan_row:
        dv_plan = DataValidation(type="list", formula1='"Yes,No,In Progress"', allow_blank=False)
        ws.add_data_validation(dv_plan)
        dv_plan.add(f"B{remediation_plan_row}")
    
    # IT OPERATIONS REVIEW
    row += 2
    ws[f"A{row}"] = "IT OPERATIONS REVIEW"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")
    
    row += 1
    itops_fields = [
        ("Reviewed By (IT Ops):", "input"),
        ("Review Date:", "input"),
        ("Technical Accuracy Status:", "dropdown"),
        ("IT Ops Comments:", "input"),
    ]
    
    itops_status_row = None
    thin = Side(style="thin")
    for label, field_type in itops_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

        if field_type == "dropdown":
            itops_status_row = row
        row += 1
    
    if itops_status_row:
        dv_itops = DataValidation(type="list", formula1='"Verified,Needs Review,Inaccurate"', allow_blank=False)
        ws.add_data_validation(dv_itops)
        dv_itops.add(f"B{itops_status_row}")
    
    # SECURITY TEAM REVIEW
    row += 2
    ws[f"A{row}"] = "SECURITY TEAM REVIEW"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")
    
    row += 1
    security_fields = [
        ("Reviewed By (Security):", "input"),
        ("Review Date:", "input"),
        ("Security Compliance Status:", "dropdown"),
        ("Security Comments:", "input"),
    ]
    
    security_status_row = None
    for label, field_type in security_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

        if field_type == "dropdown":
            security_status_row = row
        row += 1
    
    if security_status_row:
        dv_sec = DataValidation(type="list", formula1='"Compliant,Gaps Identified,Non-Compliant"', allow_blank=False)
        ws.add_data_validation(dv_sec)
        dv_sec.add(f"B{security_status_row}")
    
    # DATA PROTECTION OFFICER REVIEW
    row += 2
    ws[f"A{row}"] = "DATA PROTECTION OFFICER REVIEW"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")
    
    row += 1
    dpo_fields = [
        ("Reviewed By (DPO):", "input"),
        ("Review Date:", "input"),
        ("Data Protection Compliance:", "dropdown_compliance"),
        ("Cross-Border Transfer Status:", "dropdown_transfer"),
        ("DPO Comments:", "input"),
    ]
    
    dpo_compliance_row = None
    dpo_transfer_row = None
    for label, field_type in dpo_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

        if field_type == "dropdown_compliance":
            dpo_compliance_row = row
        elif field_type == "dropdown_transfer":
            dpo_transfer_row = row
        row += 1
    
    if dpo_compliance_row:
        dv_dpo_comp = DataValidation(type="list", formula1='"Compliant,Partially Compliant,Non-Compliant"', allow_blank=False)
        ws.add_data_validation(dv_dpo_comp)
        dv_dpo_comp.add(f"B{dpo_compliance_row}")
    
    if dpo_transfer_row:
        dv_dpo_transfer = DataValidation(type="list", formula1='"Approved,Approved with SCCs,Requires TIA,Rejected"', allow_blank=False)
        ws.add_data_validation(dv_dpo_transfer)
        dv_dpo_transfer.add(f"B{dpo_transfer_row}")
    
    # CISO APPROVAL
    row += 2
    ws[f"A{row}"] = "CISO APPROVAL"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")
    
    row += 1
    ciso_fields = [
        ("Approved By (CISO):", "input"),
        ("Approval Date:", "input"),
        ("Approval Decision:", "dropdown"),
        ("Executive Comments:", "input"),
    ]
    
    ciso_decision_row = None
    for label, field_type in ciso_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

        if field_type == "dropdown":
            ciso_decision_row = row
        row += 1

    if ciso_decision_row:
        dv_ciso = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected"', allow_blank=False)
        ws.add_data_validation(dv_ciso)
        dv_ciso.add(f"B{ciso_decision_row}")

    # Final decision
    row += 1
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin = Side(style="thin")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1
    
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 11: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution - orchestrates workbook creation."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.5.23.S3 - Secure Configuration & Deployment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.5.23: Cloud Services Security")
    logger.info("REGULATORY UPDATE: DORA, NIS2, AI Act, CLOUD Act Compliance")
    logger.info("=" * 80)
    
    wb = create_workbook()
    styles = setup_styles()
    
    logger.info("\n[1/10] Creating Instructions & Legend sheet (with regulatory guidance)...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    
    logger.info("[2/10] Creating 2. Configuration Baseline sheet (with DORA/NIS2/AI Act cols)...")
    create_2_configuration_baseline(wb["2. Configuration Baseline"], styles)
    
    logger.info("[3/10] Creating 3. Access Control Setup sheet...")
    create_3_access_control(wb["3. Access Control Setup"], styles)
    
    logger.info("[4/10] Creating 4. Network Security sheet...")
    create_4_network_security(wb["4. Network Security"], styles)
    
    logger.info("[5/10] Creating 5. Encryption Configuration sheet...")
    create_5_encryption_config(wb["5. Encryption Configuration"], styles)
    
    logger.info("[6/10] Creating 6. Deployment Checklist sheet...")
    create_6_deployment_checklist(wb["6. Deployment Checklist"], styles)
    
    logger.info("[7/10] Creating 7. Jurisdictional Risk sheet...")
    create_7_jurisdictional_risk(wb["7. Jurisdictional Risk"], styles)
    
    logger.info("[8/10] Creating 8. Summary Dashboard sheet (with regulatory metrics)...")
    create_8_summary_dashboard(wb["8. Summary Dashboard"], styles)
    
    logger.info("[9/10] Creating 9. Evidence Register sheet (expanded for regulatory)...")
    create_9_evidence_register(wb["9. Evidence Register"], styles)
    
    logger.info("[10/10] Creating 10. Approval Sign-Off sheet (with DPO section)...")
    create_10_approval_signoff(wb["10. Approval Sign-Of"], styles)
    
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.5.23.S3_SecureConfig_{timestamp}.xlsx"
    wb.save(filename)
    
    logger.info("\n" + "=" * 80)
    logger.info("{CHECK} Workbook generated successfully: {filename}")
    logger.info("=" * 80)
    logger.info("\n📋 WORKBOOK STRUCTURE (10 sheets):")
    logger.info("  1. Instructions & Legend (with regulatory applicability guidance)")
    logger.info("  2. Configuration Baseline (27 cols: base + DORA/NIS2/AI Act)")
    logger.info("  3. Access Control Setup (24 cols: SSO, MFA, RBAC, JIT)")
    logger.info("  4. Network Security (24 cols: IP restrictions, segmentation, WAF)")
    logger.info("  5. Encryption Configuration (24 cols: encryption, keys, HSM)")
    logger.info("  6. Deployment Checklist (24 cols: pre-deploy, monitoring, rollback)")
    logger.info("  8. Summary Dashboard (4 tables: compliance + jurisdictional + regulatory)")
    logger.info("  9. Evidence Register (expanded for regulatory evidence types)")
    logger.info("  10. Approval Sign-Off (IT Ops, Security, DPO, CISO)")
    
    logger.info("  • Configuration Baseline: +3 columns (DORA/NIS2/AI Act)")
    logger.info("  • Dashboard: +2 tables (jurisdictional + regulatory compliance)")
    logger.info("  • Approval: +DPO review section")
    logger.info("  • Checklists: +15 regulatory items")
    logger.info("  • Dropdowns: +9 regulatory validation types")
    
    logger.info("\n📊 REGULATORY COMPLIANCE COVERAGE:")
    logger.info("  ✓ DORA (Digital Operational Resilience Act)")
    logger.info("  ✓ NIS2 (Network and Information Security Directive 2)")
    logger.info("  ✓ EU AI Act (High-Risk AI Systems)")
    logger.info("  ✓ US CLOUD Act (Jurisdictional Risk Assessment)")
    
    logger.info("\n🔧 NEXT STEPS:")
    logger.info(f"  1. Run validation: python3 excel_sanity_check_a523.py {filename}")
    logger.info("  2. Open in Excel and verify all 10 sheets present")
    logger.info("  3. Test regulatory dropdowns (DORA/NIS2/AI Act/Jurisdictional)")
    logger.info("  4. Verify dashboard formulas calculate correctly")
    logger.info("  5. Distribute to stakeholders:")
    logger.info("     - IT Operations (configuration baseline)")
    logger.info("     - Security Team (all security controls)")
    logger.info("     - DPO (jurisdictional risk assessment)")
    logger.info("     - Legal (CLOUD Act exposure review)")
    
    logger.info("\n💡 FEYNMAN SAYS:")
    logger.info('  "The first principle is that you must not fool yourself about compliance."')
    logger.info('  Translation: Document what you ACTUALLY have, not what you WISH you had!')
    
    return filename


if __name__ == "__main__":
    main()

logger.info("Part 4 loaded: Dashboard, Evidence Register, Approval & Main Execution")
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
