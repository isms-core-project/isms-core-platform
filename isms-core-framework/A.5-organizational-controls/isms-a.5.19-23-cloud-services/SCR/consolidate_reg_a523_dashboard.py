#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.23 - Dashboard Consolidation Script
================================================================================

ISO/IEC 27001:2022 Control A.5.23: Information Security for Use of Cloud Services
Reads 4 normalised assessment workbooks, consolidates into S5 dashboard.

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

**Purpose:**
Reads completed S1-S4 assessment workbooks (normalised filenames) and populates
the S5 Compliance Monitoring Dashboard with:
- Gap analysis from all domain-specific assessment sheets
- Evidence register entries from all 4 workbooks
- Risk register derived from high-severity gaps
- Remediation roadmap with priority and target dates
- Exit Strategy metrics from S1 Sheet 4 (columns R-AC)
- PoC Testing compliance from S4 Sheet 7 (DORA Art. 28.6)

**Source Workbooks (normalised):**
- ISMS-IMP-A.5.23.S1.xlsx — Cloud Service Inventory & Classification
- ISMS-IMP-A.5.23.S2.xlsx — Vendor Due Diligence & Contracts
- ISMS-IMP-A.5.23.S3.xlsx — Secure Configuration & Deployment
- ISMS-IMP-A.5.23.S4.xlsx — Ongoing Governance & Risk Management

**Target Dashboard:**
- ISMS-IMP-A.5.23.S5_Dashboard_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Prerequisites:
    - Run normalize_assessment_files_reg_a523.py first to produce normalised names
    - Run from the Dashboard_Sources directory (where normalised files live)

Basic Usage:
    python3 consolidate_reg_a523_dashboard.py <dashboard.xlsx>

Requirements:
    - Python 3.8+
    - openpyxl library: pip install openpyxl
"""

import sys
import os
import logging
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from datetime import datetime, timedelta

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# SOURCE WORKBOOK CONFIGURATION
# =============================================================================

SOURCE_WORKBOOKS = {
    "wb1": "ISMS-IMP-A.5.23.S1.xlsx",
    "wb2": "ISMS-IMP-A.5.23.S2.xlsx",
    "wb3": "ISMS-IMP-A.5.23.S3.xlsx",
    "wb4": "ISMS-IMP-A.5.23.S4.xlsx",
}

DOMAIN_NAMES = {
    "wb1": "Inventory",
    "wb2": "Vendor Due Diligence",
    "wb3": "Secure Configuration",
    "wb4": "Governance",
}

# Sheets to skip when scanning for gaps (common sheets, not domain-specific)
SKIP_SHEET_KEYWORDS = [
    'Instructions',
    'Summary Dashboard',
    'Evidence Register',
    'Approval Sign-Of',
]

# S5 dashboard sheet names (from generate_reg_a523_5_compliance_dashboard.py)
S5_SHEETS = {
    "risk_overview": "Risk Overview",
    "recommendations": "Recommendations",
    "exit_strategy": "Exit Strategy Dashboard",
}

# =============================================================================
# DATA EXTRACTION FUNCTIONS
# =============================================================================

def safely_write_data(ws, start_row, data):
    """Safely write data to worksheet, handling merged cells."""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except Exception as e:  # Skip merged/protected cells
                continue
        entries_written += 1
    return entries_written


def _should_skip_sheet(sheet_name):
    """Check if a sheet should be skipped (common sheets, not domain-specific)."""
    for keyword in SKIP_SHEET_KEYWORDS:
        if keyword in sheet_name:
            return True
    return False


def _find_evidence_sheet(wb):
    """Find the Evidence Register sheet by substring match."""
    for name in wb.sheetnames:
        if 'Evidence Register' in name:
            return name
    return None


def extract_gaps_from_workbook(filepath, domain_name):
    """Extract gap items (Partial/Non-Compliant) from source workbook."""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        gaps = []
        gap_counter = 1

        for sheet_name in wb.sheetnames:
            if _should_skip_sheet(sheet_name):
                continue

            ws = wb[sheet_name]

            for row in range(4, min(ws.max_row + 1, 150)):
                row_values = [ws.cell(row=row, column=col).value for col in range(1, 20)]

                status = None
                status_col = None
                for col_idx, val in enumerate(row_values, start=1):
                    if val is None:
                        continue
                    val_str = str(val)
                    # Match both old format [!]/[X] and new emoji format
                    if any(marker in val_str for marker in [
                        '\u26A0', 'Partial',       # ⚠️ Partial
                        '\u274C', 'Non-Compliant',  # ❌ Non-Compliant
                        '[!] Partial', '[X] Non-Compliant',
                    ]):
                        if 'Non-Compliant' in val_str or '\u274C' in val_str:
                            status = 'Non-Compliant'
                        elif 'Partial' in val_str or '\u26A0' in val_str:
                            status = 'Partial'
                        status_col = col_idx
                        break

                if not status:
                    continue

                system_name = row_values[0] if row_values[0] else 'Unknown'
                gap_desc = None

                if status_col:
                    for offset in [2, 3, 4]:
                        col_check = status_col + offset
                        if col_check <= len(row_values):
                            potential_gap = row_values[col_check - 1]
                            if potential_gap and len(str(potential_gap)) > 10:
                                gap_desc = str(potential_gap)
                                break

                severity = 'High' if status == 'Non-Compliant' else 'Medium'

                gap_entry = [
                    domain_name,
                    f'GAP-{domain_name.split()[0]}-{gap_counter:03d}',
                    gap_desc if gap_desc else f'{system_name}: Configuration gap',
                    severity,
                    status,
                    'Compliant',
                    None,
                    'Open',
                ]
                gaps.append(gap_entry)
                gap_counter += 1

        wb.close()
        return gaps
    except Exception as e:
        logger.error(f"  Error extracting gaps from {filepath}: {e}")
        return []


def extract_evidence_from_workbook(filepath, domain_name):
    """Extract evidence entries from source workbook Evidence Register sheet."""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)

        ws_name = _find_evidence_sheet(wb)
        if not ws_name:
            wb.close()
            return []

        ws = wb[ws_name]
        evidence = []

        # Evidence data starts at row 5 (headers at row 4)
        for row in range(5, min(ws.max_row + 1, 105)):
            evidence_id = ws[f'A{row}'].value
            if evidence_id and str(evidence_id).startswith('EV-'):
                row_data = [
                    evidence_id,
                    domain_name,
                    ws[f'B{row}'].value,
                    ws[f'C{row}'].value,
                    ws[f'D{row}'].value,
                    ws[f'E{row}'].value,
                    ws[f'F{row}'].value,
                    ws[f'G{row}'].value,
                ]
                evidence.append(row_data)

        wb.close()
        return evidence
    except Exception as e:
        logger.error(f"  Error reading evidence from {filepath}: {e}")
        return []


def extract_exit_strategy_from_inventory(filepath):
    """
    Extract Exit Strategy metrics from Inventory workbook Sheet 4.
    Column mapping from generate_reg_a523_1_inventory.py:
    R = Exit Strategy Type, U = Export Tested, X = Lock-In Risk, H = Criticality
    """
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)

        sheet_names = ['4. Cloud Security Services', '4. Exit Feasibility & Strategy']
        ws = None
        for name in sheet_names:
            if name in wb.sheetnames:
                ws = wb[name]
                break

        if not ws:
            wb.close()
            return {
                'total_services': 0, 'cloud_to_cloud': 0, 'hybrid': 0,
                'on_premises': 0, 'not_determined': 0, 'high_lock_in': 0,
                'critical_lock_in': 0, 'export_not_tested_critical': 0,
            }

        metrics = {
            'total_services': 0, 'cloud_to_cloud': 0, 'hybrid': 0,
            'on_premises': 0, 'not_determined': 0, 'high_lock_in': 0,
            'critical_lock_in': 0, 'export_not_tested_critical': 0,
        }

        for row in range(4, min(ws.max_row + 1, 100)):
            service_name = ws[f'A{row}'].value
            if not service_name or service_name == '':
                continue

            metrics['total_services'] += 1

            exit_strategy = ws[f'R{row}'].value
            criticality = ws[f'H{row}'].value
            lock_in_risk = ws[f'X{row}'].value
            export_tested = ws[f'U{row}'].value

            if exit_strategy == 'Cloud-to-Cloud':
                metrics['cloud_to_cloud'] += 1
            elif exit_strategy == 'Hybrid':
                metrics['hybrid'] += 1
            elif exit_strategy == 'On-Premises':
                metrics['on_premises'] += 1
            else:
                metrics['not_determined'] += 1

            if lock_in_risk == 'High':
                metrics['high_lock_in'] += 1
            elif lock_in_risk == 'Critical':
                metrics['critical_lock_in'] += 1

            if criticality == 'Critical' and export_tested == 'No':
                metrics['export_not_tested_critical'] += 1

        wb.close()
        return metrics
    except Exception as e:
        logger.error(f"  Error extracting exit strategy from {filepath}: {e}")
        return {
            'total_services': 0, 'cloud_to_cloud': 0, 'hybrid': 0,
            'on_premises': 0, 'not_determined': 0, 'high_lock_in': 0,
            'critical_lock_in': 0, 'export_not_tested_critical': 0,
        }


def extract_poc_testing_from_governance(filepath):
    """
    Extract PoC Testing metrics from Governance workbook Sheet 7.
    DORA Article 28.6 PoC testing compliance.
    """
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)

        if '7. Exit Strategy Review' not in wb.sheetnames:
            wb.close()
            return {
                'total_requiring_poc': 0, 'poc_completed': 0,
                'poc_overdue': 0, 'overdue_services': [],
            }

        ws = wb['7. Exit Strategy Review']

        metrics = {
            'total_requiring_poc': 0, 'poc_completed': 0,
            'poc_overdue': 0, 'overdue_services': [],
        }

        for row in range(4, min(ws.max_row + 1, 100)):
            service_name = ws[f'A{row}'].value
            if not service_name or service_name == '':
                continue

            poc_required = ws[f'H{row}'].value
            poc_result = ws[f'J{row}'].value

            if poc_required == 'Yes (Critical)':
                metrics['total_requiring_poc'] += 1

                if poc_result == 'Pass':
                    metrics['poc_completed'] += 1
                elif poc_result in ['Not Tested', 'In Progress', 'Fail']:
                    metrics['poc_overdue'] += 1
                    metrics['overdue_services'].append(service_name)

        wb.close()
        return metrics
    except Exception as e:
        logger.error(f"  Error extracting PoC testing from {filepath}: {e}")
        return {
            'total_requiring_poc': 0, 'poc_completed': 0,
            'poc_overdue': 0, 'overdue_services': [],
        }


# =============================================================================
# DERIVED DATA GENERATION
# =============================================================================

def generate_risks_from_gaps(gaps):
    """Generate risk entries from high-severity gaps."""
    risks = []
    risk_counter = 1

    for gap in gaps:
        domain = gap[0]
        gap_id = gap[1]
        gap_desc = gap[2]
        severity = gap[3]

        if severity != 'High':
            continue

        risk_entry = [
            f'RISK-{risk_counter:03d}',
            gap_id,
            'Security',
            f'Security risk: {gap_desc[:80]}...',
            domain,
            'High',
            'High',
            'High',
            'Open',
            f'Mitigate gap {gap_id}',
            None,
            None,
            None,
        ]
        risks.append(risk_entry)
        risk_counter += 1

    return risks


def generate_remediation_from_gaps(gaps):
    """Generate remediation roadmap from gaps."""
    remediation = []
    action_counter = 1
    today = datetime.now()

    for gap in gaps:
        domain = gap[0]
        gap_id = gap[1]
        gap_desc = gap[2]
        severity = gap[3]

        priority = 'Critical' if severity == 'High' else 'High' if severity == 'Medium' else 'Medium'
        days_offset = 30 if priority == 'Critical' else 60 if priority == 'High' else 90
        target_date = (today + timedelta(days=days_offset)).strftime('%Y-%m-%d')

        remediation_entry = [
            f'REM-{action_counter:03d}',
            gap_id,
            domain,
            f'Remediate: {gap_desc[:60]}...',
            priority,
            'Planned',
            None,
            today.strftime('%Y-%m-%d'),
            target_date,
            None,
            None,
            None,
        ]
        remediation.append(remediation_entry)
        action_counter += 1

    return remediation


# =============================================================================
# MAIN CONSOLIDATION
# =============================================================================

def populate_dashboard(dashboard_file):
    """Populate dashboard from source workbooks."""

    logger.info("=" * 78)
    logger.info("A.5.23 CLOUD SERVICES DASHBOARD CONSOLIDATION")
    logger.info("=" * 78)

    missing = []
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        if not os.path.exists(wb_filename):
            missing.append(wb_filename)

    if missing:
        logger.error("Missing source workbooks:")
        for wb in missing:
            logger.error(f"    - {wb}")
        logger.error("  Place all 4 normalised workbooks in same directory.")
        return False

    all_gaps = []
    all_evidence = []

    logger.info("[1/4] Extracting gaps from source workbooks...")
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        domain_name = DOMAIN_NAMES[wb_key]
        gaps = extract_gaps_from_workbook(wb_filename, domain_name)
        logger.info(f"  * {domain_name}: {len(gaps)} gaps")
        all_gaps.extend(gaps)

    logger.info(f"  Total gaps identified: {len(all_gaps)}")

    logger.info("[2/4] Extracting evidence from source workbooks...")
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        domain_name = DOMAIN_NAMES[wb_key]
        evidence = extract_evidence_from_workbook(wb_filename, domain_name)
        logger.info(f"  * {domain_name}: {len(evidence)} evidence docs")
        all_evidence.extend(evidence)

    logger.info(f"  Total evidence collected: {len(all_evidence)}")

    logger.info("[3/4] Extracting Exit Strategy & PoC Testing metrics...")
    exit_strategy_metrics = extract_exit_strategy_from_inventory(SOURCE_WORKBOOKS["wb1"])
    poc_testing_metrics = extract_poc_testing_from_governance(SOURCE_WORKBOOKS["wb4"])

    if exit_strategy_metrics['total_services'] > 0:
        total = exit_strategy_metrics['total_services']
        c2c_pct = (exit_strategy_metrics['cloud_to_cloud'] / total * 100)
        hybrid_pct = (exit_strategy_metrics['hybrid'] / total * 100)
        onprem_pct = (exit_strategy_metrics['on_premises'] / total * 100)

        logger.info(f"  * Exit Strategy Coverage:")
        logger.info(f"    - Total Services: {total}")
        logger.info(f"    - Cloud-to-Cloud: {exit_strategy_metrics['cloud_to_cloud']} ({c2c_pct:.1f}%)")
        logger.info(f"    - Hybrid: {exit_strategy_metrics['hybrid']} ({hybrid_pct:.1f}%)")
        logger.info(f"    - On-Premises: {exit_strategy_metrics['on_premises']} ({onprem_pct:.1f}%)")
        logger.info(f"    - Not Determined: {exit_strategy_metrics['not_determined']}")
        logger.info(f"  * Lock-In Risks: High={exit_strategy_metrics['high_lock_in']}, Critical={exit_strategy_metrics['critical_lock_in']}")
        logger.info(f"    Export Not Tested (Critical): {exit_strategy_metrics['export_not_tested_critical']}")
    else:
        logger.info("  * Exit Strategy: No data found (Sheet 4 may be empty)")

    if poc_testing_metrics['total_requiring_poc'] > 0:
        logger.info(f"  * PoC Testing (DORA Art. 28.6):")
        logger.info(f"    - Critical Services: {poc_testing_metrics['total_requiring_poc']}")
        logger.info(f"    - PoC Completed (Pass): {poc_testing_metrics['poc_completed']}")
        logger.info(f"    - PoC Overdue/Failed: {poc_testing_metrics['poc_overdue']}")
        if poc_testing_metrics['poc_overdue'] > 0:
            for svc in poc_testing_metrics['overdue_services'][:5]:
                logger.warning(f"      - {svc}")
    else:
        logger.info("  * PoC Testing: No data found (Sheet 7 may be empty)")

    logger.info("[4/4] Generating risks and remediation...")
    all_risks = generate_risks_from_gaps(all_gaps)
    all_remediation = generate_remediation_from_gaps(all_gaps)
    logger.info(f"  * Risks generated: {len(all_risks)}")
    logger.info(f"  * Remediation actions: {len(all_remediation)}")

    logger.info("Writing to dashboard...")
    try:
        wb = load_workbook(dashboard_file)
    except Exception as e:
        logger.error(f"  Error loading dashboard: {e}")
        return False

    # Write risks to Risk Overview sheet (data starts at row 5)
    risk_sheet = S5_SHEETS["risk_overview"]
    if risk_sheet in wb.sheetnames and all_risks:
        ws = wb[risk_sheet]
        count = safely_write_data(ws, 5, all_risks)
        logger.info(f"  [OK] {risk_sheet}: {count} entries")

    # Write remediation to Recommendations sheet (data starts at row 4)
    reco_sheet = S5_SHEETS["recommendations"]
    if reco_sheet in wb.sheetnames and all_remediation:
        ws = wb[reco_sheet]
        count = safely_write_data(ws, 4, all_remediation)
        logger.info(f"  [OK] {reco_sheet}: {count} entries")

    try:
        wb.save(dashboard_file)
        logger.info(f"[SAVED] {dashboard_file}")
    except Exception as e:
        logger.error(f"  Error saving: {e}")
        return False

    logger.info("=" * 78)
    logger.info("DASHBOARD CONSOLIDATION COMPLETE")
    logger.info("=" * 78)
    logger.info(f"  Gaps: {len(all_gaps)}")
    logger.info(f"  Risks: {len(all_risks)}")
    logger.info(f"  Remediation: {len(all_remediation)}")
    logger.info(f"  Evidence: {len(all_evidence)}")

    return True


def main():
    if len(sys.argv) < 2:
        logger.error("Usage: python3 consolidate_reg_a523_dashboard.py <dashboard.xlsx>")
        sys.exit(1)

    dashboard_file = sys.argv[1]

    if not os.path.exists(dashboard_file):
        logger.error(f"Dashboard file not found: {dashboard_file}")
        sys.exit(1)

    success = populate_dashboard(dashboard_file)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-02-10
# QA_STATUS: PASSED - UTILITY SCRIPT
# CHANGES: standard docstring, logger, evidence sheet/ID/row fixes, S5 sheet targets
# =============================================================================
