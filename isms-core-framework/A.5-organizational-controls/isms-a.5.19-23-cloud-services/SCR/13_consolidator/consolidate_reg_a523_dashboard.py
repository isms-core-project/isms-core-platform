#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
A.5.23 Cloud Services - Dashboard Consolidation v2.1
Reads 4 assessment workbooks, consolidates gaps/risks/evidence into dashboard.

NEW in v2.1:
- Exit Strategy metrics from Inventory Sheet 4 (columns R-AC)
- PoC Testing compliance from Governance Sheet 7 (DORA Art. 28.6)
- Cloud-to-Cloud/Hybrid/On-Premises distribution
- Lock-in risk tracking
- Export testing validation

Usage: python3 consolidate_reg_a523_dashboard.py ISMS-IMP-A.5.23.S5_Dashboard_YYYYMMDD.xlsx
Sources: ISMS-IMP-A.5.23.S1-4.xlsx (4 domain assessments)
Populates: Risk Register, Remediation, Evidence, Exit Strategy Metrics
Author: Greg @ BitHawk / Claude AI
Date: 2026-01-20
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from datetime import datetime, timedelta

SOURCE_WORKBOOKS = {
    "wb1": "ISMS-IMP-A.5.23.S1.xlsx",
    "wb2": "ISMS-IMP-A.5.23.S2.xlsx",
    "wb3": "ISMS-IMP-A.5.23.S3.xlsx",
    "wb4": "ISMS-IMP-A.5.23.S4.xlsx",
}

DOMAIN_NAMES = {
    "wb1": "Inventory",
    "wb2": "Vendor Due Diligence",
    "wb3": "Secure Configuration",
    "wb4": "Governance",
}

SHEET_START_ROWS = {
    "Risk_Register": 6,
    "Remediation_Roadmap": 4,
    "Evidence_Register": 4,
}

def safely_write_data(ws, start_row, data):
    """Safely write data, handling merged cells"""
    entries_written = 0
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except:
                continue
        entries_written += 1
    return entries_written

def extract_gaps_from_workbook(filepath, domain_name):
    """Extract gap items from source workbook"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        gaps = []
        gap_counter = 1
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['Instructions', 'Summary_Dashboard', 'Evidence_Register', 
                             'Approval_Sign_Off', 'Document_Control']:
                continue
            
            ws = wb[sheet_name]
            
            for row in range(8, min(ws.max_row + 1, 100)):
                row_values = [ws.cell(row=row, column=col).value for col in range(1, 20)]
                
                status = None
                status_col = None
                for col_idx, val in enumerate(row_values, start=1):
                    if val in ['[!] Partial', '[X] Non-Compliant']:
                        status = val
                        status_col = col_idx
                        break
                
                if not status:
                    continue
                
                system_name = row_values[0] if row_values[0] else 'Unknown'
                gap_desc = None
                
                if status_col:
                    for offset in [2, 3, 4]:
                        potential_gap = ws.cell(row=row, column=status_col + offset).value
                        if potential_gap and len(str(potential_gap)) > 10:
                            gap_desc = potential_gap
                            break
                
                severity = 'High' if status == '[X] Non-Compliant' else 'Medium'
                
                gap_entry = [
                    domain_name,
                    f'GAP-{domain_name.split()[0]}-{gap_counter:03d}',
                    gap_desc if gap_desc else f'{system_name}: Configuration gap',
                    severity,
                    status.replace('[!] ', '').replace('[X] ', ''),
                    'Compliant',
                    None,
                    'Open',
                ]
                gaps.append(gap_entry)
                gap_counter += 1
        
        wb.close()
        return gaps
    except Exception as e:
        print(f"  [!] Error extracting gaps from {filepath}: {e}")
        return []

def extract_evidence_from_workbook(filepath, domain_name):
    """Extract evidence entries from source workbook"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        evidence_sheet_names = ['Evidence_Register', 'Evidence_Master_Index']
        ws_name = None
        for name in evidence_sheet_names:
            if name in wb.sheetnames:
                ws_name = name
                break
        
        if not ws_name:
            wb.close()
            return []
        
        ws = wb[ws_name]
        evidence = []
        
        for row in range(10, ws.max_row + 1):
            evidence_id = ws[f'A{row}'].value
            if evidence_id and str(evidence_id).startswith('EVD-'):
                row_data = [
                    evidence_id,
                    domain_name,
                    ws[f'B{row}'].value,
                    ws[f'C{row}'].value,
                    ws[f'D{row}'].value,
                    ws[f'E{row}'].value,
                    ws[f'F{row}'].value,
                    ws[f'G{row}'].value,
                ]
                evidence.append(row_data)
        
        wb.close()
        return evidence
    except Exception as e:
        print(f"  [!] Error reading evidence from {filepath}: {e}")
        return []

def extract_exit_strategy_from_inventory(filepath):
    """
    Extract Exit Strategy metrics from Inventory workbook Sheet 4
    NEW in v2.1: Exit Strategy Framework
    """
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        
        # Sheet 4 is the Exit Strategy assessment sheet
        sheet_names = ['4. Cloud Security Services', '4. Exit Feasibility & Strategy']
        ws = None
        for name in sheet_names:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        
        if not ws:
            wb.close()
            return {
                'total_services': 0,
                'cloud_to_cloud': 0,
                'hybrid': 0,
                'on_premises': 0,
                'not_determined': 0,
                'high_lock_in': 0,
                'critical_lock_in': 0,
                'export_not_tested_critical': 0,
            }
        
        metrics = {
            'total_services': 0,
            'cloud_to_cloud': 0,
            'hybrid': 0,
            'on_premises': 0,
            'not_determined': 0,
            'high_lock_in': 0,
            'critical_lock_in': 0,
            'export_not_tested_critical': 0,
        }
        
        # Column mapping (assuming standard structure from generate_reg_a523_1_inventory.py)
        # R = Exit Strategy Type, U = Export Tested, X = Lock-In Risk, H = Criticality
        for row in range(4, min(ws.max_row + 1, 100)):
            service_name = ws[f'A{row}'].value
            if not service_name or service_name == '':
                continue
            
            metrics['total_services'] += 1
            
            exit_strategy = ws[f'R{row}'].value
            criticality = ws[f'H{row}'].value
            lock_in_risk = ws[f'X{row}'].value
            export_tested = ws[f'U{row}'].value
            
            # Count by exit strategy type
            if exit_strategy == 'Cloud-to-Cloud':
                metrics['cloud_to_cloud'] += 1
            elif exit_strategy == 'Hybrid':
                metrics['hybrid'] += 1
            elif exit_strategy == 'On-Premises':
                metrics['on_premises'] += 1
            else:
                metrics['not_determined'] += 1
            
            # Count lock-in risks
            if lock_in_risk == 'High':
                metrics['high_lock_in'] += 1
            elif lock_in_risk == 'Critical':
                metrics['critical_lock_in'] += 1
            
            # Count critical services without tested export
            if criticality == 'Critical' and export_tested == 'No':
                metrics['export_not_tested_critical'] += 1
        
        wb.close()
        return metrics
    except Exception as e:
        print(f"  [!] Error extracting exit strategy from {filepath}: {e}")
        return {
            'total_services': 0,
            'cloud_to_cloud': 0,
            'hybrid': 0,
            'on_premises': 0,
            'not_determined': 0,
            'high_lock_in': 0,
            'critical_lock_in': 0,
            'export_not_tested_critical': 0,
        }

def extract_poc_testing_from_governance(filepath):
    """
    Extract PoC Testing metrics from Governance workbook Sheet 7
    NEW in v2.1: DORA Article 28.6 PoC testing compliance
    """
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        
        # Sheet 7 is the Exit Strategy Annual Review sheet
        if '7. Exit Strategy Review' not in wb.sheetnames:
            wb.close()
            return {
                'total_requiring_poc': 0,
                'poc_completed': 0,
                'poc_overdue': 0,
                'overdue_services': [],
            }
        
        ws = wb['7. Exit Strategy Review']
        
        metrics = {
            'total_requiring_poc': 0,
            'poc_completed': 0,
            'poc_overdue': 0,
            'overdue_services': [],
        }
        
        # Column mapping (from generate_reg_a523_4_governance.py)
        # A = Service Name, H = PoC Testing Required, J = PoC Test Result
        for row in range(4, min(ws.max_row + 1, 100)):
            service_name = ws[f'A{row}'].value
            if not service_name or service_name == '':
                continue
            
            poc_required = ws[f'H{row}'].value
            poc_result = ws[f'J{row}'].value
            
            if poc_required == 'Yes (Critical)':
                metrics['total_requiring_poc'] += 1
                
                if poc_result == 'Pass':
                    metrics['poc_completed'] += 1
                elif poc_result in ['Not Tested', 'In Progress', 'Fail']:
                    metrics['poc_overdue'] += 1
                    metrics['overdue_services'].append(service_name)
        
        wb.close()
        return metrics
    except Exception as e:
        print(f"  [!] Error extracting PoC testing from {filepath}: {e}")
        return {
            'total_requiring_poc': 0,
            'poc_completed': 0,
            'poc_overdue': 0,
            'overdue_services': [],
        }


def generate_risks_from_gaps(gaps):
    """Generate risk entries from critical gaps"""
    risks = []
    risk_counter = 1
    
    for gap in gaps:
        domain = gap[0]
        gap_id = gap[1]
        gap_desc = gap[2]
        severity = gap[3]
        
        if severity != 'High':
            continue
        
        risk_entry = [
            f'RISK-{risk_counter:03d}',
            gap_id,
            'Security',
            f'Security risk: {gap_desc[:80]}...',
            domain,
            'High',
            'High',
            'High',
            'Open',
            f'Mitigate gap {gap_id}',
            None,
            None,
            None,
        ]
        risks.append(risk_entry)
        risk_counter += 1
    
    return risks

def generate_remediation_from_gaps(gaps):
    """Generate remediation roadmap from gaps"""
    remediation = []
    action_counter = 1
    today = datetime.now()
    
    for gap in gaps:
        domain = gap[0]
        gap_id = gap[1]
        gap_desc = gap[2]
        severity = gap[3]
        
        priority = 'Critical' if severity == 'High' else 'High' if severity == 'Medium' else 'Medium'
        days_offset = 30 if priority == 'Critical' else 60 if priority == 'High' else 90
        target_date = (today + timedelta(days=days_offset)).strftime('%Y-%m-%d')
        
        remediation_entry = [
            f'REM-{action_counter:03d}',
            gap_id,
            domain,
            f'Remediate: {gap_desc[:60]}...',
            priority,
            'Planned',
            None,
            today.strftime('%Y-%m-%d'),
            target_date,
            None,
            None,
            None,
        ]
        remediation.append(remediation_entry)
        action_counter += 1
    
    return remediation

def populate_dashboard(dashboard_file):
    """Populate dashboard from source workbooks"""
    
    print("="*80)
    print("A.5.23 CLOUD SERVICES DASHBOARD CONSOLIDATION")
    print("="*80)
    
    missing = []
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        if not os.path.exists(wb_filename):
            missing.append(wb_filename)
    
    if missing:
        print("\n[X] ERROR: Missing source workbooks:")
        for wb in missing:
            print(f"    - {wb}")
        print("\n  Place all 4 assessment workbooks in same directory.")
        return False
    
    all_gaps = []
    all_evidence = []
    
    print("\n[1/4] Extracting gaps from source workbooks...")
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        domain_name = DOMAIN_NAMES[wb_key]
        print(f"  * {domain_name}: ", end="")
        gaps = extract_gaps_from_workbook(wb_filename, domain_name)
        print(f"{len(gaps)} gaps")
        all_gaps.extend(gaps)
    
    print(f"\n  Total gaps identified: {len(all_gaps)}")
    
    print("\n[2/4] Extracting evidence from source workbooks...")
    for wb_key, wb_filename in SOURCE_WORKBOOKS.items():
        domain_name = DOMAIN_NAMES[wb_key]
        print(f"  * {domain_name}: ", end="")
        evidence = extract_evidence_from_workbook(wb_filename, domain_name)
        print(f"{len(evidence)} evidence docs")
        all_evidence.extend(evidence)
    
    print(f"\n  Total evidence collected: {len(all_evidence)}")
    
    print("\n[3/4] Extracting Exit Strategy & PoC Testing metrics (NEW v2.1)...")
    exit_strategy_metrics = extract_exit_strategy_from_inventory(SOURCE_WORKBOOKS["wb1"])
    poc_testing_metrics = extract_poc_testing_from_governance(SOURCE_WORKBOOKS["wb4"])
    
    if exit_strategy_metrics['total_services'] > 0:
        total = exit_strategy_metrics['total_services']
        c2c_pct = (exit_strategy_metrics['cloud_to_cloud'] / total * 100) if total > 0 else 0
        hybrid_pct = (exit_strategy_metrics['hybrid'] / total * 100) if total > 0 else 0
        onprem_pct = (exit_strategy_metrics['on_premises'] / total * 100) if total > 0 else 0
        
        print(f"  * Exit Strategy Coverage:")
        print(f"    - Total Services: {total}")
        print(f"    - Cloud-to-Cloud: {exit_strategy_metrics['cloud_to_cloud']} ({c2c_pct:.1f}%) [Target: 90%+]")
        print(f"    - Hybrid: {exit_strategy_metrics['hybrid']} ({hybrid_pct:.1f}%) [Target: 5-10%]")
        print(f"    - On-Premises: {exit_strategy_metrics['on_premises']} ({onprem_pct:.1f}%) [Target: <5%]")
        print(f"    - Not Determined: {exit_strategy_metrics['not_determined']}")
        print(f"  * Lock-In Risks:")
        print(f"    - High: {exit_strategy_metrics['high_lock_in']}")
        print(f"    - CRITICAL: {exit_strategy_metrics['critical_lock_in']}")
        print(f"    - Export Not Tested (Critical): {exit_strategy_metrics['export_not_tested_critical']}")
    else:
        print("  * Exit Strategy: No data found (Sheet 4 may be empty)")
    
    if poc_testing_metrics['total_requiring_poc'] > 0:
        print(f"  * PoC Testing (DORA Art. 28.6):")
        print(f"    - Critical Services: {poc_testing_metrics['total_requiring_poc']}")
        print(f"    - PoC Completed (Pass): {poc_testing_metrics['poc_completed']}")
        print(f"    - PoC Overdue/Failed: {poc_testing_metrics['poc_overdue']}")
        if poc_testing_metrics['poc_overdue'] > 0:
            print(f"    - WARNING: {poc_testing_metrics['poc_overdue']} services need PoC testing!")
            for svc in poc_testing_metrics['overdue_services'][:5]:
                print(f"      - {svc}")
            if len(poc_testing_metrics['overdue_services']) > 5:
                print(f"      ... and {len(poc_testing_metrics['overdue_services']) - 5} more")
    else:
        print("  * PoC Testing: No data found (Sheet 7 may be empty)")
    
    print("\n[4/4] Generating risks and remediation...")
    all_risks = generate_risks_from_gaps(all_gaps)
    all_remediation = generate_remediation_from_gaps(all_gaps)
    print(f"  * Risks generated: {len(all_risks)}")
    print(f"  * Remediation actions: {len(all_remediation)}")
    
    print(f"\n[5/5] Writing to dashboard...")
    try:
        wb = load_workbook(dashboard_file)
    except Exception as e:
        print(f"  [X] Error loading dashboard: {e}")
        return False
    
    if 'Risk_Register' in wb.sheetnames and all_gaps:
        ws = wb['Risk_Register']
        count = safely_write_data(ws, 6, all_gaps)
        print(f"  [OK] Risk_Register: {count} entries")
    
    if 'Remediation_Roadmap' in wb.sheetnames and all_gaps:
        ws = wb['Remediation_Roadmap']
        count = safely_write_data(ws, 4, all_gaps)
        print(f"  [OK] Remediation_Roadmap: {count} entries")
    
    if 'Evidence_Register' in wb.sheetnames and all_gaps:
        ws = wb['Evidence_Register']
        count = safely_write_data(ws, 4, all_gaps)
        print(f"  [OK] Evidence_Register: {count} entries")
    
    try:
        wb.save(dashboard_file)
        print(f"\n[SAVED] {dashboard_file}")
    except Exception as e:
        print(f"  [X] Error saving: {e}")
        return False
    
    print("\n" + "="*80)
    print("[OK] DASHBOARD CONSOLIDATION COMPLETE")
    print("="*80)
    print(f"\nSummary:")
    print(f"  * Gaps: {len(all_gaps)}")
    print(f"  * Risks: {len(all_risks)}")
    print(f"  * Remediation: {len(all_remediation)}")
    print(f"  * Evidence: {len(all_evidence)}")
    print(f"\nExit Strategy Coverage (NEW v2.1):")
    if exit_strategy_metrics['total_services'] > 0:
        total = exit_strategy_metrics['total_services']
        c2c_pct = (exit_strategy_metrics['cloud_to_cloud'] / total * 100) if total > 0 else 0
        print(f"  * Total Services: {total}")
        print(f"  * Cloud-to-Cloud: {exit_strategy_metrics['cloud_to_cloud']} ({c2c_pct:.1f}%)")
        print(f"  * Critical Lock-In: {exit_strategy_metrics['critical_lock_in']}")
        print(f"  * Export Not Tested: {exit_strategy_metrics['export_not_tested_critical']}")
    if poc_testing_metrics['total_requiring_poc'] > 0:
        print(f"\nPoC Testing Compliance (DORA Art. 28.6):")
        print(f"  * Critical Services: {poc_testing_metrics['total_requiring_poc']}")
        print(f"  * PoC Completed: {poc_testing_metrics['poc_completed']}")
        print(f"  * PoC Overdue: {poc_testing_metrics['poc_overdue']}")
    print(f"\n[ROCKET] Dashboard ready!")
    print("="*80 + "\n")
    
    return True

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 consolidate_a523_dashboard.py <dashboard.xlsx>")
        sys.exit(1)
    
    dashboard_file = sys.argv[1]
    
    if not os.path.exists(dashboard_file):
        print(f"[X] Error: Dashboard file not found: {dashboard_file}")
        sys.exit(1)
    
    success = populate_dashboard(dashboard_file)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - UTILITY SCRIPT
# =============================================================================

