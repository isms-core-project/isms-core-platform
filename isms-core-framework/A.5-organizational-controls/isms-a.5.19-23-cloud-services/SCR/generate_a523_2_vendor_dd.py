#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.23.S2 - Vendor Due Diligence & Contracts Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.23: Information Security for Use of Cloud Services
Assessment Domain 2 of 5: Vendor Due Diligence & Contracts

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific vendor management processes, contract requirements,
and due diligence criteria.

Key customization areas:
1. Due diligence criteria (security certifications per your requirements)
2. Contract requirements (SLA terms, data protection clauses)
3. Regulatory requirements (DORA, NIS2 per your jurisdictions)
4. Risk assessment methodology (based on your risk framework)
5. Approval workflows (per your procurement governance)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework (adapted for cloud services)

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

**Purpose:**
Enables systematic vendor due diligence and contract review for cloud services
supporting ISO 27001:2022 Control A.5.23 requirements and regulatory compliance.

**Assessment Scope:**
- Vendor security assessment (certifications, audits)
- Contract security clause verification
- Data protection agreement review
- Exit strategy and data portability
- Subprocessor management
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance
2. Vendor Assessment - Due diligence evaluation
3. Contract Review - Security clause checklist
4. Data Protection - DPA and privacy requirements
5. Exit Strategy - Portability and termination rights
6. Subprocessor Management - Third-party tracking
7. Gap Analysis - Non-compliant vendors/contracts
8. Evidence Register - Audit evidence tracking
9. Approval & Sign-Off - Stakeholder approval

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a523_2_vendor_dd.py

Requirements:
    - Python 3.8+
    - openpyxl library: pip install openpyxl

Output:
    ISMS-IMP-A.5.23.S2_VendorDD_YYYYMMDD.xlsx
"""

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.23.S2"
WORKBOOK_NAME = "Vendor Due Diligence & Contracts"
CONTROL_ID = "A.5.19-23"
CONTROL_NAME = "Information Security for Use of Cloud Services"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Row configuration
MAX_DATA_ROWS = 50  # Standard maximum data rows per DS-005

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# REGULATORY DROPDOWN CONSTANTS (DORA, NIS2, AI Act, CLOUD Act)
# ============================================================================

PROVIDER_HQ_JURISDICTION = [
    "Switzerland", "EU/EEA", "United Kingdom", "United States",
    "Other Adequate Country", "Non-Adequate Country"
]

CLOUD_ACT_EXPOSURE = [
    "No Exposure", "Potential Exposure (US HQ)", "Mitigated (EU Data Boundary)",
    "Mitigated (Encryption + Key Control)", "Accepted Risk (Documented)", "Under Assessment"
]

TRANSFER_MECHANISM = ["SCCs", "BCRs", "Adequacy Decision", "None", "N/A"]
RISK_LEVEL = ["Low", "Medium", "High", "Critical"]
YES_NO_PARTIAL_UNKNOWN = ["Yes", "No", "Partial", "Unknown"]
YES_NO_PLANNED = ["Yes", "No", "Planned"]

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠ Warning sign
CLOUD = '\u2601'      # ☁ Cloud
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets (10 sheets incl. Jurisdictional Risk)."""
    wb = Workbook()
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Sheet structure - v1.0 includes Jurisdictional Risk
    sheets = [
        "Instructions & Legend",
        "2. Security Certifications",
        "3. Contract Terms",
        "4. SLA Performance",
        "5. Data Sovereignty",
        "6. Forensics & Audit",
        "7. Jurisdictional Risk",
        "8. Summary Dashboard",         # Was 7
        "9. Evidence Register",         # Was 8
        "10. Approval Sign-Of",        # Was 9
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    
    return wb


def setup_styles() -> dict:
    """Define all cell styles. CRITICAL: Create NEW objects per cell!"""
    return {
        "header": {
            "font_params": {"name": "Calibri", "size": 14, "bold": True, "color": "FFFFFF"},
            "fill_params": {"start_color": "003366", "end_color": "003366", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "subheader": {
            "font_params": {"name": "Calibri", "size": 11, "bold": True, "color": "FFFFFF"},
            "fill_params": {"start_color": "4472C4", "end_color": "4472C4", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "column_header": {
            "font_params": {"name": "Calibri", "size": 10, "bold": True},
            "fill_params": {"start_color": "D9D9D9", "end_color": "D9D9D9", "fill_type": "solid"},
            "alignment_params": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        },
        "input_cell": {
            "fill_params": {"start_color": "FFFFCC", "end_color": "FFFFCC", "fill_type": "solid"},
            "alignment_params": {"horizontal": "left", "vertical": "center", "wrap_text": True},
        },
    }


def apply_style(cell, style_def, style_type):
    """Apply style to cell using NEW objects (avoids shared object issues)."""
    if "font_params" in style_def:
        cell.font = Font(**style_def["font_params"])
    if "fill_params" in style_def:
        cell.fill = PatternFill(**style_def["fill_params"])
    if "alignment_params" in style_def:
        cell.alignment = Alignment(**style_def["alignment_params"])
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: COLUMN DEFINITIONS (BASE + EXTENDED)
# ============================================================================

def get_base_vendor_columns() -> dict:
    """Return base 17 columns (A-Q) standard for ALL vendor DD assessments."""
    return {
        "Cloud Service Name": 30, "Vendor Name": 25, "Service Type": 22,
        "Service Criticality": 20, "Data Classification": 22, "Contract Type": 22,
        "Contract Start Date": 18, "Status": 15, "Evidence Location": 30,
        "Gap Description": 35, "Remediation Needed": 16, "Exception ID": 14,
        "Risk ID": 14, "Compensating Controls": 35, "Vendor Contact (Legal)": 25,
        "Target Remediation Date": 18, "Contract Owner": 22,
    }


def get_extended_columns_certifications() -> dict:
    """Extended columns R-X for security certifications."""
    return {
        "ISO 27001 Certified": 18, "ISO 27001 Cert Number": 22, "ISO 27001 Expiry Date": 18,
        "SOC 2 Type II Report": 20, "SOC 2 Report Date": 18, "FedRAMP Authorised": 18,
        "Other Certifications": 30,
    }


def get_extended_columns_contracts() -> dict:
    """Extended columns R-X for contract terms."""
    return {
        "Data Protection Clause": 20, "Subprocessor Disclosure": 20,
        "Liability Cap (CHF)": 20, "Indemnification Clause": 18,
        "Termination Notice Period": 20, "Data Return on Termination": 20,
        "Auto-Renewal Clause": 18,

        "DORA Art.30 Compliance": 20, "NIS2 Supply Chain Clause": 20,
    }


def get_extended_columns_sla() -> dict:
    """Extended columns R-X for SLA performance."""
    return {
        "SLA Availability %": 18, "Actual Availability Q4": 20, "SLA Credits Claimed": 18,
        "Support Response Time": 22, "Incident Count Q4": 16, "Mean Time to Resolve": 18,
        "SLA Review Date": 18,
    }


def get_extended_columns_sovereignty() -> dict:
    """Extended columns R-X for data sovereignty."""
    return {
        "Data Processing Region": 22, "Data Residency Verified": 20,
        "Standard Contractual Clauses": 22, "Data Transfer Mechanism": 22,
        "GDPR Compliance": 18, "Swiss nFADP Compliance": 20, "Regulatory Framework": 30,
    }


def get_extended_columns_forensics() -> dict:
    """Extended columns R-X for forensics & audit."""
    return {
        "Forensics Support": 20, "Forensics SLA": 20, "Right to Audit": 18,
        "Audit Notice Period": 18, "Incident Notification SLA": 22,
        "Breach Notification": 20, "IR Playbook Provided": 18,
    }


def get_jurisdictional_columns() -> dict:
    """Columns for NEW Jurisdictional Risk Assessment sheet."""
    return {
        "Assessment ID": 14, "Cloud Service Name": 25, "Provider Name": 22,
        "Provider HQ Country": 18, "Provider HQ Jurisdiction": 20,
        "US Parent Company": 14, "CLOUD Act Applicability": 20,
        "Data Processing Locations": 25, "EU Data Boundary Available": 18,
        "Customer Managed Keys": 16, "Legal Challenge Commitment": 18,
        "Adequacy Decision Status": 18, "Transfer Mechanism": 16,
        "Risk Level": 14, "Risk Accepted By": 18, "Risk Acceptance Date": 16,
        "Compensating Controls": 28, "Review Date": 14, "Evidence Reference": 20,
        "Notes": 30,
    }


def get_all_columns(sheet_type: str) -> dict:
    """Combine base + extended columns based on sheet type."""
    base = get_base_vendor_columns()
    
    extended_map = {
        "certifications": get_extended_columns_certifications,
        "contracts": get_extended_columns_contracts,
        "sla": get_extended_columns_sla,
        "sovereignty": get_extended_columns_sovereignty,
        "forensics": get_extended_columns_forensics,
    }
    
    extended_func = extended_map.get(sheet_type)
    if extended_func:
        return {**base, **extended_func()}
    return base


# ============================================================================
# SECTION 3: VALIDATION DEFINITIONS
# ============================================================================

def create_base_validations(ws) -> dict:
    """Create data validation objects for base columns."""
    validations = {}
    
    validations['service_type'] = DataValidation(
        type="list", formula1='"SaaS,IaaS,PaaS,Security Service,Cloud Storage,Other"', allow_blank=False)
    validations['criticality'] = DataValidation(
        type="list", formula1='"Critical,High,Medium,Low"', allow_blank=False)
    validations['data_classification'] = DataValidation(
        type="list", formula1='"Public,Internal,Confidential,Restricted,Mixed"', allow_blank=False)
    validations['contract_type'] = DataValidation(
        type="list", formula1='"MSA + DPA,Subscription Agreement,Pay-As-You-Go,Trial,Custom"', allow_blank=False)
    validations['status'] = DataValidation(
        type="list", formula1=f'"{CHECK} Compliant,⚠️ Partial,❌ Non-Compliant,N/A"', allow_blank=False)
    validations['yes_no'] = DataValidation(
        type="list", formula1='"Yes,No"', allow_blank=False)
    
    return validations


def finalize_validations(ws, validations):
    """Add only data validations that have cells assigned to avoid Excel repair."""
    for dv in validations.values():
        if dv.sqref:
            ws.add_data_validation(dv)


def create_extended_validations_certifications(ws) -> dict:
    """Validations for security certifications extended columns."""
    v = {}
    v['iso_certified'] = DataValidation(type="list", formula1='"Yes (Current),Yes (Expired),No,Unknown"', allow_blank=False)
    v['soc2_report'] = DataValidation(type="list", formula1='"Yes (< 6 months),Yes (6-12 months),No,Unknown"', allow_blank=False)
    v['fedramp'] = DataValidation(type="list", formula1='"Yes (High),Yes (Moderate),Yes (Low),No,N/A"', allow_blank=False)
    return v


def create_extended_validations_contracts(ws) -> dict:
    """Validations for contract terms extended columns."""
    v = {}
    v['data_protection'] = DataValidation(type="list", formula1='"Yes (Adequate),Yes (Weak),No,Under Negotiation"', allow_blank=False)
    v['subprocessor'] = DataValidation(type="list", formula1='"Yes (List Provided),Yes (Generic),No"', allow_blank=False)
    v['indemnification'] = DataValidation(type="list", formula1='"Yes (Favorable),Yes (Limited),No"', allow_blank=False)
    v['termination_notice'] = DataValidation(type="list", formula1='"≤30 days,31-60 days,61-90 days,>90 days"', allow_blank=False)
    v['data_return'] = DataValidation(type="list", formula1='"Yes (30 days),Yes (60 days),Yes (90 days),No"', allow_blank=False)
    v['auto_renewal'] = DataValidation(type="list", formula1='"Yes (Opt-Out),Yes (Opt-In),No"', allow_blank=False)

    v['dora_compliance'] = DataValidation(type="list", formula1='"Yes (Full),Yes (Partial),No,N/A (Not in scope)"', allow_blank=False)
    v['nis2_clause'] = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=False)
    return v


def create_jurisdictional_validations(ws) -> dict:
    """Validations for NEW Jurisdictional Risk Assessment sheet."""
    v = {}
    v['hq_jurisdiction'] = DataValidation(type="list", formula1='"' + ','.join(PROVIDER_HQ_JURISDICTION) + '"', allow_blank=False)
    v['cloud_act'] = DataValidation(type="list", formula1='"' + ','.join(CLOUD_ACT_EXPOSURE) + '"', allow_blank=False)
    v['yes_no_unknown'] = DataValidation(type="list", formula1='"Yes,No,Unknown"', allow_blank=False)
    v['yes_no_partial'] = DataValidation(type="list", formula1='"' + ','.join(YES_NO_PARTIAL_UNKNOWN) + '"', allow_blank=False)
    v['yes_no_planned'] = DataValidation(type="list", formula1='"' + ','.join(YES_NO_PLANNED) + '"', allow_blank=False)
    v['transfer'] = DataValidation(type="list", formula1='"' + ','.join(TRANSFER_MECHANISM) + '"', allow_blank=False)
    v['risk_level'] = DataValidation(type="list", formula1='"' + ','.join(RISK_LEVEL) + '"', allow_blank=False)
    return v


# --- END BLOCK 1 ---
logger.info("Block 1 loaded: Foundation (Sections 1-3)")

# ============================================================================
# SECTION 4: CHECKLIST DEFINITIONS
# ============================================================================

def get_checklist_certifications() -> list:
    """Return checklist items for security certifications."""
    return [
        ("All critical/high services have ISO 27001 or SOC 2 Type II", "Certification register"),
        ("Certificates verified as current (not expired)", "Certificate expiry tracking"),
        ("SOC 2 reports obtained within last 12 months", "Vendor portal, NDA required"),
        ("Certificate scope covers services in use", "Certificate scope statement"),
        ("Certificate issuing body is accredited (e.g., UKAS, ANAB)", "Issuer verification"),
        ("FedRAMP authorization verified for US government workloads", "FedRAMP marketplace"),
        ("Industry-specific certs documented (PCI-DSS, HIPAA, etc.)", "Compliance attestations"),
        ("Certificate renewal dates tracked in contract management system", "CMS reports"),
    ]


def get_checklist_contracts() -> list:
    """Return checklist items for contract terms - UPDATED with DORA/NIS2."""
    base_items = [
        ("Data protection clause references GDPR/nFADP", "Contract review"),
        ("Subprocessor list reviewed by Legal", "Legal sign-of"),
        ("Liability cap reasonable for service criticality", "Risk assessment"),
        ("Termination notice ≤90 days for non-critical services", "Contract terms"),
        ("Data return process documented with timeline", "Exit annex"),
        ("Insurance requirements verified (cyber liability)", "Insurance cert"),
        ("Governing law and jurisdiction acceptable", "Legal review"),
    ]
    # DORA Article 30 contract clauses (for financial sector)
    dora_items = [
        ("DORA Art.30(2)(a): Clear service descriptions", "Contract clause"),
        ("DORA Art.30(2)(b): Data processing locations specified", "DPA/Contract"),
        ("DORA Art.30(2)(c): Subcontracting conditions defined", "Subprocessor annex"),
        ("DORA Art.30(2)(d): Full access and audit rights", "Audit clause"),
        ("DORA Art.30(2)(e): Cooperation with regulators", "Regulatory clause"),
        ("DORA Art.30(3): Exit strategy provisions", "Exit annex"),
        ("DORA Art.30(3): Termination rights defined", "Termination clause"),
    ]
    # NIS2 supply chain security clauses
    nis2_items = [
        ("NIS2 Art.21: Supply chain security requirements", "Security annex"),
        ("NIS2: Incident notification ≤24h to customer", "Incident clause"),
        ("NIS2: Vulnerability handling process defined", "Security annex"),
        ("Supplier incident notification to customer within 24 hours (contractual SLA documented)", "Contract security annex, SLA document"),
        ("Right to conduct security assessments and penetration testing of supplier infrastructure (contractual right confirmed)", "Audit rights clause, penetration test agreement"),
    ]
    return base_items + dora_items + nis2_items


def get_checklist_sla() -> list:
    """Return checklist items for SLA performance."""
    return [
        ("SLA availability target documented (e.g., 99.9%)", "SLA document"),
        ("SLA credits/penalties defined for breaches", "SLA terms"),
        ("Support response times defined by severity", "Support matrix"),
        ("Incident escalation path documented", "Support procedures"),
        ("Planned maintenance notification requirements", "SLA terms"),
        ("Performance metrics reporting frequency defined", "SLA reporting"),
        ("SLA review cadence established (quarterly/annual)", "Contract terms"),
        ("Historical SLA performance reviewed", "Vendor reports"),
    ]


def get_checklist_sovereignty() -> list:
    """Return checklist items for data sovereignty."""
    return [
        ("Data processing locations documented", "DPA, vendor docs"),
        ("Data residency meets regulatory requirements", "Compliance review"),
        ("Standard Contractual Clauses (SCCs) in place", "DPA annex"),
        ("Transfer Impact Assessment completed", "TIA document"),
        ("GDPR compliance verified", "Compliance attestation"),
        ("Swiss nFADP compliance verified", "Compliance attestation"),
        ("Restricted data not processed outside approved regions", "Data mapping"),
        ("Sub-processor locations reviewed and approved", "Sub-processor list"),
        ("Swiss nFADP Transfer Impact Assessment completed for transfers to non-adequate countries (beyond EU adequacy)", "TIA document, legal counsel sign-off"),
    ]


def get_checklist_forensics() -> list:
    """Return checklist items for forensics & audit rights."""
    return [
        ("Forensics support verified for critical/high services", "Vendor capability"),
        ("Forensics SLA documented (log retention, data export)", "IR annex"),
        ("Right to audit clause in contract", "Contract clause"),
        ("Audit notice period ≤30 days for critical services", "Audit terms"),
        ("Incident notification SLA ≤24 hours", "Incident SLA"),
        ("Breach notification meets GDPR 72-hour requirement", "Breach clause"),
        ("Incident response playbook obtained and reviewed", "IR playbook"),
        ("Vendor participates in tabletop exercises (if critical)", "Exercise records"),
    ]


def get_checklist_jurisdictional() -> list:
    """Return checklist items for NEW Jurisdictional Risk Assessment."""
    return [
        ("Provider HQ jurisdiction identified and documented", "Vendor documentation"),
        ("US parent company status verified", "Corporate structure review"),
        ("CLOUD Act exposure assessed for US-nexus providers", "Legal review"),
        ("Data processing locations documented in DPA", "DPA review"),
        ("EU Data Boundary availability checked", "Vendor announcement/docs"),
        ("Customer-managed encryption keys option evaluated", "Technical documentation"),
        ("Legal challenge commitment verified", "Contract/Public statement"),
        ("Transfer mechanism documented (SCCs, BCRs, etc.)", "DPA"),
        ("Risk acceptance recorded if residual risk remains", "Risk register"),
        ("Compensating controls documented for high-risk providers", "Control documentation"),
    ]


def get_regulatory_eval_criteria() -> list:
    """Regulatory evaluation criteria (EVAL-REG and EVAL-AI) - for reference."""
    return [
        ("EVAL-REG-01", "Provider discloses HQ jurisdiction and legal entity structure"),
        ("EVAL-REG-02", "Provider discloses all data processing locations"),
        ("EVAL-REG-03", "Provider offers EU Data Boundary or regional commitment"),
        ("EVAL-REG-04", "Provider supports customer-managed encryption keys"),
        ("EVAL-REG-05", "Provider commits to challenge government data requests"),
        ("EVAL-REG-06", "Provider discloses sub-processor jurisdictions"),
        ("EVAL-AI-01", "AI system risk classification documented (per EU AI Act)"),
        ("EVAL-AI-02", "High-risk AI system has conformity assessment"),
        ("EVAL-AI-03", "AI transparency requirements met"),
    ]


# ============================================================================
# SECTION 5: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws):
    """Create Instructions & Legend sheet with regulatory compliance info."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header — single merged row with two-line title
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.5.23.S2  -  Vendor Due Diligence & Contracts\n"
        "ISO/IEC 27001:2022 - Control A.5.23: Information Security for Use of Cloud Services"
    )
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].border = border
    ws.row_dimensions[1].height = 40

    # Document Information
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.5.23.S2"),
        ("Assessment Area", "Vendor Due Diligence & Contracts"),
        ("Related Policy", "ISMS-POL-A.5.19-23-S1, S2"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = border
        row += 1

    # Regulatory Compliance Section
    row += 1
    ws[f"A{row}"] = "REGULATORY COMPLIANCE"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="003366")

    row += 1
    reg_info = [
        "This workbook includes regulatory compliance tracking for:",
        f"{BULLET} DORA (Digital Operational Resilience Act) - Art.30 contract clauses",
        f"{BULLET} NIS2 (Network and Information Security Directive 2) - Supply chain security",
        f"{BULLET} EU AI Act - AI service provider evaluation criteria",
        f"{BULLET} US CLOUD Act - Jurisdictional risk assessment",
        "",
        "Sheet 7 (Jurisdictional Risk) helps assess CLOUD Act exposure for US-nexus providers.",
        "Contract Terms checklist now includes DORA Art.30 and NIS2 requirements.",
    ]
    for line in reg_info:
        ws[f"A{row}"] = line
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        row += 1

    # How to Use This Workbook
    row += 1
    ws[f"A{row}"] = "Instructions"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Complete each worksheet tab (2-7) for all cloud service vendors.",
        "2. Use dropdown menus for standardised entries.",
        "3. Fill in yellow-highlighted cells with vendor-specific information.",
        "4. Attach evidence: contracts (MSA, DPA, SLA), security certifications, audit reports.",
        "5. Complete Jurisdictional Risk sheet (7) for all US-nexus providers.",
        "6. Document SLA performance metrics and track contract renewal dates.",
        "7. Update vendor security posture assessments annually or upon significant changes.",
        "8. Summary Dashboard auto-calculates vendor compliance statistics.",
        "9. Maintain the Evidence Register with all contract documents and certifications.",
        "10. Obtain final approval and sign-off from Legal, Procurement, and CISO.",
    ]
    row += 1
    for instr in instructions:
        ws[f"A{row}"] = instr
        row += 1

    # Status Legend — proper table with headers
    row += 1
    ws[f"A{row}"] = "Status Legend"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    for ci, hdr in enumerate(("Symbol", "Status", "Description"), start=1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    legend = [
        (CHECK, "Compliant", "Vendor meets all security and contractual requirements", "C6EFCE"),
        (WARNING, "Partial", "Some requirements met, gaps exist (e.g., cert pending)", "FFEB9C"),
        (XMARK, "Non-Compliant", "Vendor does not meet minimum requirements", "FFC7CE"),
        ("—", "Not Applicable", "Requirement does not apply to this vendor/service", None),
    ]

    row += 1
    for sym, status, desc, color in legend:
        ws.cell(row=row, column=1, value=sym).border = border
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if color:
            s.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

    # Acceptable Evidence
    row += 1
    ws[f"A{row}"] = "ACCEPTABLE EVIDENCE (examples)"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    evidence_types = [
        "✓ Master Service Agreement (MSA)",
        "✓ Data Processing Agreement (DPA) / Addendum",
        "✓ Service Level Agreement (SLA)",
        "✓ ISO 27001 certificate (current, not expired)",
        "✓ SOC 2 Type II report (within last 12 months)",
        "✓ FedRAMP authorization letter",
        "✓ GDPR/nFADP compliance attestation",
        "✓ Vendor security questionnaire (completed)",
        "✓ Penetration test results (if available)",
        "✓ Incident response plan documentation",
        "✓ Business continuity/disaster recovery plan",
        "✓ Insurance certificates (cyber liability)",
        "✓ Right to audit clause (contract section reference)",
        "✓ Data sovereignty documentation (data center locations)",
        "✓ Jurisdictional risk assessment (for US-nexus providers)",
    ]
    row += 1
    for ev in evidence_types:
        ws[f"A{row}"] = ev
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 6: GENERIC ASSESSMENT SHEET ENGINE
# ============================================================================

def create_vendor_assessment_sheet(ws, styles, section_title, policy_req, 
                                   question, sheet_type, checklist_items):
    """
    Generic assessment sheet creator for vendor due diligence.
    
    Args:
        ws: worksheet object
        styles: style dictionary
        section_title: e.g., "VENDOR SECURITY CERTIFICATIONS"
        policy_req: Policy reference text
        question: Assessment question
        sheet_type: 'certifications', 'contracts', 'sla', 'sovereignty', 'forensics'
        checklist_items: List of (requirement, evidence) tuples
    """
    # Get columns for this sheet type
    all_columns = get_all_columns(sheet_type)
    col_names = list(all_columns.keys())
    col_widths = list(all_columns.values())
    last_col = get_column_letter(len(col_names))
    
    # Section Header
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"{section_title}\n{policy_req}"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 50
    
    # Assessment Question
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Assessment Question: {question}"
    apply_style(ws["A2"], styles["subheader"], "subheader")
    ws.row_dimensions[2].height = 30
    
    # Column Headers (Row 3)
    row = 3
    for col_idx, (col_name, col_width) in enumerate(zip(col_names, col_widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        apply_style(cell, styles["column_header"], "header")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # Empty data rows (rows 5-54: 50 FFFFCC rows = 51 total with F2F2F2 sample at row 4)
    for data_row in range(5, 55):
        for col_idx in range(1, len(col_names) + 1):
            cell = ws.cell(row=data_row, column=col_idx, value="")
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Sample row (row 4) — F2F2F2 grey with realistic example data
    thin_s = Side(style="thin")
    border_s = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
    _sample_values = {
        "certifications": ["CERT-001", "Vendor Security Certifications", "Amazon Web Services", "Critical",
                           "Confidential", "Annual", "SaaS/IaaS", "Verified", "AWS Security Team",
                           "ISO 27001:2022 + SOC 2 Type II + CSA STAR", "", "", "", "", "", "", ""],
        "contracts": ["CONTRACT-001", "Cloud Service Contract Review", "Microsoft Azure", "Critical",
                      "Confidential", "Annual", "SaaS/PaaS", "Reviewed", "Legal Team",
                      "Full contract + DPA + SLA", "", "", "", "", "", "", ""],
        "sla": ["SLA-001", "SLA Performance Review", "AWS", "Critical", "Internal",
                "Monthly", "IaaS", "Within SLA", "IT Operations",
                "SLA performance dashboard export", "", "", "", "", "", "", ""],
        "sovereignty": ["SOV-001", "Data Sovereignty Assessment", "Microsoft Azure", "Critical",
                        "Confidential", "Annual", "SaaS", "Compliant", "DPO",
                        "DPA + Data Residency Confirmation", "", "", "", "", "", "", ""],
        "forensics": ["FOR-001", "Forensics & Audit Access Review", "AWS", "High", "Internal",
                      "Annual", "IaaS", "Compliant", "Security Team",
                      "Audit agreement + IR playbook", "", "", "", "", "", "", ""],
    }
    _row4_vals = _sample_values.get(sheet_type, [])
    for _ci in range(1, len(col_names) + 1):
        _cell = ws.cell(row=4, column=_ci, value=_row4_vals[_ci - 1] if _ci <= len(_row4_vals) else "")
        _cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        _cell.font = Font(italic=True, color="666666")
        _cell.border = border_s

    # Apply base validations
    base_vals = create_base_validations(ws)
    base_vals['service_type'].add("C4:C54")
    base_vals['criticality'].add("D4:D54")
    base_vals['data_classification'].add("E4:E54")
    base_vals['contract_type'].add("F4:F54")
    base_vals['status'].add("H4:H54")
    base_vals['yes_no'].add("K4:K54")

    # Apply sheet-specific extended validations
    ext_vals = None
    if sheet_type == "certifications":
        ext_vals = create_extended_validations_certifications(ws)
        ext_vals['iso_certified'].add("R4:R54")
        ext_vals['soc2_report'].add("U4:U54")
        ext_vals['fedramp'].add("W4:W54")
    elif sheet_type == "contracts":
        ext_vals = create_extended_validations_contracts(ws)
        ext_vals['data_protection'].add("R4:R54")
        ext_vals['subprocessor'].add("S4:S54")
        ext_vals['indemnification'].add("U4:U54")
        ext_vals['termination_notice'].add("V4:V54")
        ext_vals['data_return'].add("W4:W54")
        ext_vals['auto_renewal'].add("X4:X54")
        ext_vals['dora_compliance'].add("Y4:Y54")
        ext_vals['nis2_clause'].add("Z4:Z54")
    elif sheet_type == "sovereignty":
        _create_sovereignty_validations(ws)
    elif sheet_type == "forensics":
        _create_forensics_validations(ws)

    # Finalize: add only validations that have cells assigned
    finalize_validations(ws, base_vals)
    if ext_vals:
        finalize_validations(ws, ext_vals)
    
    # Checklist Section (starts after data rows end at 54)
    if checklist_items:
        checklist_row = 57
        ws[f"A{checklist_row}"] = "COMPLIANCE CHECKLIST"
        ws[f"A{checklist_row}"].font = Font(bold=True, size=11)
        ws.merge_cells(f"A{checklist_row}:D{checklist_row}")
        
        checklist_row += 1
        headers = ["☐", "Requirement", "Status", "Evidence Type"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=checklist_row, column=col_idx, value=h)
            apply_style(cell, styles["column_header"], "header")
        
        # Status validation for checklist
        status_dv = DataValidation(type="list", formula1=f'"{CHECK},"{WARNING}","{XMARK}",N/A"', allow_blank=True)
        ws.add_data_validation(status_dv)
        
        checklist_row += 1
        for req, evidence in checklist_items:
            ws.cell(row=checklist_row, column=1, value="☐")
            ws.cell(row=checklist_row, column=2, value=req)
            status_cell = ws.cell(row=checklist_row, column=3, value="")
            status_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            status_cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            status_dv.add(status_cell)
            ws.cell(row=checklist_row, column=4, value=evidence)
            checklist_row += 1

    ws.freeze_panes = "A4"


def _create_sovereignty_validations(ws):
    """Create and apply Data Sovereignty-specific validations."""
    dv_region = DataValidation(type="list", formula1='"Switzerland,EU/EEA,USA,Asia-Pacific,Global,Unknown"', allow_blank=False)
    ws.add_data_validation(dv_region)
    dv_region.add("R4:R54")

    dv_verified = DataValidation(type="list", formula1='"Yes (Contractual),Yes (Technical),No,Unknown"', allow_blank=False)
    ws.add_data_validation(dv_verified)
    dv_verified.add("S4:S54")

    dv_scc = DataValidation(type="list", formula1='"Yes (EU SCC),Yes (Swiss SCC),Yes (Other),No,N/A"', allow_blank=False)
    ws.add_data_validation(dv_scc)
    dv_scc.add("T4:T54")

    dv_transfer = DataValidation(type="list", formula1='"Adequacy Decision,SCC,BCR,Derogations,None"', allow_blank=False)
    ws.add_data_validation(dv_transfer)
    dv_transfer.add("U4:U54")

    dv_gdpr = DataValidation(type="list", formula1='"Yes (Certified),Yes (Self-Assessed),No,N/A"', allow_blank=False)
    ws.add_data_validation(dv_gdpr)
    dv_gdpr.add("V4:V54")
    dv_gdpr.add("W4:W54")  # Also for nFADP


def _create_forensics_validations(ws):
    """Create and apply Forensics & Audit-specific validations."""
    dv_forensics = DataValidation(type="list", formula1='"Yes (Full),Yes (Limited),No,Unknown"', allow_blank=False)
    ws.add_data_validation(dv_forensics)
    dv_forensics.add("R4:R54")

    dv_audit = DataValidation(type="list", formula1='"Yes (Unrestricted),Yes (With Notice),No"', allow_blank=False)
    ws.add_data_validation(dv_audit)
    dv_audit.add("T4:T54")

    dv_notice = DataValidation(type="list", formula1='"No Notice,≤7 days,8-30 days,>30 days,N/A"', allow_blank=False)
    ws.add_data_validation(dv_notice)
    dv_notice.add("U4:U54")

    dv_incident = DataValidation(type="list", formula1='"<1 hour,1-4 hours,4-24 hours,>24 hours,None"', allow_blank=False)
    ws.add_data_validation(dv_incident)
    dv_incident.add("V4:V54")

    dv_breach = DataValidation(type="list", formula1='"Yes (72 hours),Yes (Custom),No,Unknown"', allow_blank=False)
    ws.add_data_validation(dv_breach)
    dv_breach.add("W4:W54")

    dv_playbook = DataValidation(type="list", formula1='"Yes,No,Upon Request"', allow_blank=False)
    ws.add_data_validation(dv_playbook)
    dv_playbook.add("X4:X54")


# --- END BLOCK 2 ---
logger.info("Block 2 loaded: Checklists, Instructions, Generic Engine (Sections 4-6)")

# ============================================================================
# SECTION 7: INDIVIDUAL SHEET CREATORS (Sheets 2-6)
# ============================================================================

def create_2_security_certifications(ws, styles):
    """Sheet 2: Vendor Security Certifications."""
    create_vendor_assessment_sheet(
        ws=ws, styles=styles,
        section_title="VENDOR SECURITY CERTIFICATIONS",
        policy_req="Policy Requirement: All cloud vendors must maintain ISO 27001 or equivalent security certification (Policy S2 Section 3.1)",
        question="Do your cloud service vendors hold current security certifications (ISO 27001, SOC 2, FedRAMP, etc.)?",
        sheet_type="certifications",
        checklist_items=get_checklist_certifications()
    )


def create_3_contract_terms(ws, styles):
    """Sheet 3: Contract Terms Analysis."""
    create_vendor_assessment_sheet(
        ws=ws, styles=styles,
        section_title="CONTRACT TERMS ANALYSIS (INCL. DORA/NIS2)",
        policy_req="Policy Requirement: All cloud service contracts must include security clauses, data protection terms, and termination rights (Policy S2 Section 3.2)",
        question="Have all cloud service contracts been reviewed for security, data protection, and regulatory clauses (DORA Art.30, NIS2)?",
        sheet_type="contracts",
        checklist_items=get_checklist_contracts()
    )


def create_4_sla_performance(ws, styles):
    """Sheet 4: SLA Requirements & Performance."""
    create_vendor_assessment_sheet(
        ws=ws, styles=styles,
        section_title="SLA REQUIREMENTS & PERFORMANCE TRACKING",
        policy_req="Policy Requirement: SLAs must define availability, support response times, and performance penalties (Policy S2 Section 3.3)",
        question="Do all cloud services have documented SLAs with performance commitments?",
        sheet_type="sla",
        checklist_items=get_checklist_sla()
    )


def create_5_data_sovereignty(ws, styles):
    """Sheet 5: Data Sovereignty & Compliance."""
    create_vendor_assessment_sheet(
        ws=ws, styles=styles,
        section_title="DATA SOVEREIGNTY & COMPLIANCE",
        policy_req="Policy Requirement: Vendor data processing locations must comply with organisational data residency requirements (Policy S2 Section 3.4)",
        question="Do vendors comply with data sovereignty and regulatory requirements for your data?",
        sheet_type="sovereignty",
        checklist_items=get_checklist_sovereignty()
    )


def create_6_forensics_audit(ws, styles):
    """Sheet 6: Forensics, Audit Rights & Incident Response."""
    create_vendor_assessment_sheet(
        ws=ws, styles=styles,
        section_title="FORENSICS, AUDIT RIGHTS & INCIDENT RESPONSE",
        policy_req="Policy Requirement: Vendors must support forensic investigations and grant audit rights (Policy S2 Section 3.5)",
        question="Do vendors provide forensic capabilities, audit rights, and incident response support?",
        sheet_type="forensics",
        checklist_items=get_checklist_forensics()
    )


# ============================================================================
# SECTION 8: JURISDICTIONAL RISK ASSESSMENT SHEET (CLOUD Act)
# ============================================================================

def create_7_jurisdictional_risk(ws, styles):
    """
    Sheet 7: NEW Jurisdictional Risk Assessment for CLOUD Act analysis.
    
    This sheet assesses jurisdictional risks for cloud providers, particularly
    US-headquartered providers subject to the CLOUD Act (Clarifying Lawful 
    Overseas Use of Data Act).
    """
    columns = get_jurisdictional_columns()
    col_names = list(columns.keys())
    col_widths = list(columns.values())
    last_col = get_column_letter(len(col_names))
    
    # Header
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "JURISDICTIONAL RISK ASSESSMENT\nCLOUD Act, Data Sovereignty, and Cross-Border Transfer Analysis"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 50
    
    # Subheader with guidance
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = "Policy Requirement: Assess jurisdictional risks for all cloud providers, particularly US-headquartered providers subject to CLOUD Act (Policy S5 Section 4)"
    apply_style(ws["A2"], styles["subheader"], "subheader")
    ws.row_dimensions[2].height = 30
    
    # Guidance text
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = f"{WARNING} US-headquartered providers (or providers with US parent companies) may be compelled to disclose data under CLOUD Act regardless of data location. Assess and document mitigations."
    ws["A3"].font = Font(italic=True, size=10, color="C00000")
    ws["A3"].alignment = Alignment(wrap_text=True)
    
    # Column headers (row 5)
    header_row = 5
    for col_idx, (col_name, col_width) in enumerate(zip(col_names, col_widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        apply_style(cell, styles["column_header"], "header")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # Data entry rows (rows 7-57: 1 sample + 50 empty = 51 total)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    start_row, end_row = 7, 57
    for row in range(start_row, end_row + 1):
        # FFFFCC input cells for ALL columns (including ID column)
        for col_idx in range(1, len(col_names) + 1):
            cell = ws.cell(row=row, column=col_idx, value="")
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = border
    
    # Create and apply validations
    v = create_jurisdictional_validations(ws)
    
    # Apply validations to correct columns (by column index) - updated to 51 rows
    # Col 5: Provider HQ Jurisdiction
    v['hq_jurisdiction'].add("E7:E57")
    # Col 6: US Parent Company
    v['yes_no_unknown'].add("F7:F57")
    # Col 7: CLOUD Act Applicability
    v['cloud_act'].add("G7:G57")
    # Col 9: EU Data Boundary Available
    v['yes_no_planned'].add("I7:I57")
    # Col 10: Customer Managed Keys
    v['yes_no_planned'].add("J7:J57")
    # Col 11: Legal Challenge Commitment
    v['yes_no_partial'].add("K7:K57")
    # Col 13: Transfer Mechanism
    v['transfer'].add("M7:M57")
    # Col 14: Risk Level
    v['risk_level'].add("N7:N57")

    # Example row (row 7) - Microsoft 365 as reference
    example_data = [
        ("A7", "JRA-001"),
        ("B7", "Microsoft 365"),
        ("C7", "Microsoft Corporation"),
        ("D7", "United States"),
        ("E7", "United States"),
        ("F7", "No"),
        ("G7", "Mitigated (EU Data Boundary)"),
        ("H7", "EU, Switzerland (EU Data Boundary enabled)"),
        ("I7", "Yes"),
        ("J7", "Yes"),
        ("K7", "Yes"),
        ("L7", "Adequacy Decision (Swiss-US DPF)"),
        ("M7", "SCCs"),
        ("N7", "Medium"),
        ("O7", "CISO"),
        ("P7", "2025-01-15"),
        ("Q7", "EU Data Boundary + Customer Lockbox"),
        ("R7", "2025-07-01"),
        ("S7", "EV-JRA-001"),
        ("T7", "Example entry - delete or overwrite"),
    ]
    for cell_ref, value in example_data:
        ws[cell_ref] = value
        ws[cell_ref].font = Font(italic=True, color="666666")
    
    # Checklist Section (starts after data rows end at 57)
    checklist = get_checklist_jurisdictional()
    checklist_row = 60
    
    ws[f"A{checklist_row}"] = "JURISDICTIONAL RISK CHECKLIST"
    ws[f"A{checklist_row}"].font = Font(bold=True, size=11)
    ws.merge_cells(f"A{checklist_row}:D{checklist_row}")
    
    checklist_row += 1
    headers = ["☐", "Requirement", "Status", "Evidence Type"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=checklist_row, column=col_idx, value=h)
        apply_style(cell, styles["column_header"], "header")
    
    # Status validation for checklist
    status_dv = DataValidation(type="list", formula1=f'"{CHECK},"{WARNING}","{XMARK}",N/A"', allow_blank=True)
    ws.add_data_validation(status_dv)
    
    checklist_row += 1
    for req, evidence in checklist:
        ws.cell(row=checklist_row, column=1, value="☐")
        ws.cell(row=checklist_row, column=2, value=req)
        status_cell = ws.cell(row=checklist_row, column=3, value="")
        status_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin = Side(style="thin")
        status_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        status_dv.add(status_cell)
        ws.cell(row=checklist_row, column=4, value=evidence)
        checklist_row += 1
    
    # CLOUD Act Mitigation Strategies Reference
    checklist_row += 2
    ws[f"A{checklist_row}"] = "CLOUD ACT MITIGATION STRATEGIES (Reference)"
    ws[f"A{checklist_row}"].font = Font(bold=True, size=11)
    
    mitigations = [
        ("Technical", "EU Data Boundary, Customer-managed encryption keys, Confidential Computing"),
        ("Contractual", "Legal challenge commitment, Data disclosure notification, SCCs with supplementary measures"),
        ("Organisational", "Risk acceptance by CISO/DPO, Documented risk assessment, Regular review cycle"),
        ("Alternative", "Consider EU-headquartered alternatives for highest-risk data categories"),
    ]
    
    checklist_row += 1
    for category, description in mitigations:
        ws[f"A{checklist_row}"] = f"{BULLET} {category}:"
        ws[f"A{checklist_row}"].font = Font(bold=True)
        ws[f"B{checklist_row}"] = description
        ws.merge_cells(f"B{checklist_row}:E{checklist_row}")
        checklist_row += 1

    # Finalize: add only validations that have cells assigned
    finalize_validations(ws, v)

    ws.freeze_panes = "A6"


# ============================================================================
# SECTION 8 (continued): SUMMARY DASHBOARD
# ============================================================================

def create_8_summary_dashboard(ws, styles):
    """Sheet 8: Summary Dashboard with compliance metrics."""
    
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "VENDOR DUE DILIGENCE - COMPLIANCE SUMMARY DASHBOARD"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 35
    
    # TABLE 1: Compliance Summary by Assessment Area
    row = 3
    ws[f"A{row}"] = "TABLE 1: COMPLIANCE SUMMARY BY ASSESSMENT AREA"
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    
    row += 1
    headers = ["Assessment Area", "Total Vendors", f"{CHECK} Compliant", f"{WARNING} Partial", f"{XMARK} Non-Compliant", "N/A", "Compliance %"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Assessment areas with formula references
    areas = [
        ("Security Certifications", "'2. Security Certifications'", "H"),
        ("Contract Terms", "'3. Contract Terms'", "H"),
        ("SLA Performance", "'4. SLA Performance'", "H"),
        ("Data Sovereignty", "'5. Data Sovereignty'", "H"),
        ("Forensics & Audit", "'6. Forensics & Audit'", "H"),
        ("Jurisdictional Risk", "'7. Jurisdictional Risk'", "N"),
    ]
    
    row += 1
    start_formula_row = row
    for label, sheet, status_col in areas:
        ws.cell(row=row, column=1, value=label)
        # Updated ranges: vendor assessment sheets to row 80 (max checklist end), Jurisdictional Risk to row 71
        if label == "Jurisdictional Risk":
            ws.cell(row=row, column=2, value=f"=COUNTA({sheet}!A7:A71)-COUNTBLANK({sheet}!B7:B71)")
            ws.cell(row=row, column=3, value=f'=COUNTIF({sheet}!{status_col}7:{status_col}71,"Low")+COUNTIF({sheet}!{status_col}7:{status_col}71,"Medium")')
            ws.cell(row=row, column=4, value=f'=COUNTIF({sheet}!{status_col}7:{status_col}71,"High")')
            ws.cell(row=row, column=5, value=f'=COUNTIF({sheet}!{status_col}7:{status_col}71,"Critical")')
            ws.cell(row=row, column=6, value="=0")
        else:
            ws.cell(row=row, column=2, value=f"=COUNTA({sheet}!A5:A80)-COUNTBLANK({sheet}!B5:B80)")
            ws.cell(row=row, column=3, value=f'=COUNTIF({sheet}!{status_col}5:{status_col}80,"{CHECK}*")')
            ws.cell(row=row, column=4, value=f'=COUNTIF({sheet}!{status_col}5:{status_col}80,"{WARNING}*")')
            ws.cell(row=row, column=5, value=f'=COUNTIF({sheet}!{status_col}5:{status_col}80,"{XMARK}*")')
            ws.cell(row=row, column=6, value=f'=COUNTIF({sheet}!{status_col}5:{status_col}80,"N/A")')

        ws.cell(row=row, column=7, value=f'=IF(B{row}=0,"N/A",ROUND((C{row}/(B{row}-F{row}))*100,1)&"%")')
        row += 1
    
    # Total compliance row
    end_formula_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL")
    ws[f"A{row}"].font = Font(bold=True)
    for col in range(2, 7):
        ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_formula_row}:{get_column_letter(col)}{end_formula_row})")
    ws.cell(row=row, column=7, value=f'=IF(B{row}=0,"N/A",ROUND((C{row}/(B{row}-F{row}))*100,1)&"%")')
    
    # TABLE 2: NEW - JURISDICTIONAL & CLOUD ACT RISK SUMMARY
    row += 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "TABLE 2: JURISDICTIONAL & CLOUD ACT RISK SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    
    row += 2
    for col_idx, h in enumerate(["Metric", "Count", "Status"], start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row += 1
    # FML-004: Extend range from 7:30 to 7:57 to cover all data rows
    jur_metrics = [
        ("US-HQ Providers (CLOUD Act Scope)",
         '=COUNTIF(\'7. Jurisdictional Risk\'!E7:E71,"United States")'),
        ("Providers with US Parent Company",
         '=COUNTIF(\'7. Jurisdictional Risk\'!F7:F71,"Yes")'),
        ("CLOUD Act Potential Exposure (Unmitigated)",
         '=COUNTIF(\'7. Jurisdictional Risk\'!G7:G71,"Potential*")'),
        ("CLOUD Act Mitigated",
         '=COUNTIF(\'7. Jurisdictional Risk\'!G7:G71,"Mitigated*")'),
        ("High/Critical Jurisdictional Risk",
         '=COUNTIF(\'7. Jurisdictional Risk\'!N7:N71,"High")+COUNTIF(\'7. Jurisdictional Risk\'!N7:N71,"Critical")'),
        ("Providers Without EU Data Boundary",
         '=COUNTIF(\'7. Jurisdictional Risk\'!I7:I71,"No")'),
        ("Providers Without Customer-Managed Keys",
         '=COUNTIF(\'7. Jurisdictional Risk\'!J7:J71,"No")'),
    ]
    
    for metric, formula in jur_metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=3, value=f'=IF(B{row}>0,"{WARNING} Review Required","{CHECK} OK")')
        row += 1
    
    # TABLE 3: Security Certification Status
    row += 2
    ws[f"A{row}"] = "TABLE 3: SECURITY CERTIFICATION STATUS"
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")

    row += 1
    cert_headers = ["Certification", "Yes (Current)", "Expired/Pending", "No/Unknown"]
    for col_idx, h in enumerate(cert_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    row += 1
    ws.cell(row=row, column=1, value="ISO 27001")
    ws.cell(row=row, column=2, value="=COUNTIF('2. Security Certifications'!R5:R67,\"Yes (Current)\")")
    ws.cell(row=row, column=3, value="=COUNTIF('2. Security Certifications'!R5:R67,\"Yes (Expired)\")")
    ws.cell(row=row, column=4, value="=COUNTIF('2. Security Certifications'!R5:R67,\"No\")+COUNTIF('2. Security Certifications'!R5:R67,\"Unknown\")")
    
    row += 1
    ws.cell(row=row, column=1, value="SOC 2 Type II")
    ws.cell(row=row, column=2, value="=COUNTIF('2. Security Certifications'!U5:U67,\"Yes*\")")
    ws.cell(row=row, column=3, value="=0")
    ws.cell(row=row, column=4, value="=COUNTIF('2. Security Certifications'!U5:U67,\"No\")+COUNTIF('2. Security Certifications'!U5:U67,\"Unknown\")")
    
    row += 1
    ws.cell(row=row, column=1, value="FedRAMP")
    ws.cell(row=row, column=2, value="=COUNTIF('2. Security Certifications'!W5:W67,\"Yes*\")")
    ws.cell(row=row, column=3, value="=0")
    ws.cell(row=row, column=4, value="=COUNTIF('2. Security Certifications'!W5:W67,\"No\")+COUNTIF('2. Security Certifications'!W5:W67,\"N/A\")")
    
    # TABLE 4: Data Sovereignty Risks
    row += 3
    ws[f"A{row}"] = "TABLE 4: DATA SOVEREIGNTY RISKS"
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    
    row += 1
    risk_headers = ["Risk", "Count", "Severity"]
    for col_idx, h in enumerate(risk_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    risks = [
        ("Restricted Data Outside Switzerland/EU", "=COUNTIF('5. Data Sovereignty'!R5:R67,\"USA\")+COUNTIF('5. Data Sovereignty'!R5:R67,\"Asia-Pacific\")", "Critical"),
        ("Confidential Data Without SCCs", "=COUNTIF('5. Data Sovereignty'!T5:T67,\"No\")", "High"),
        ("GDPR Compliance Not Verified", "=COUNTIF('5. Data Sovereignty'!V5:V67,\"No\")+COUNTIF('5. Data Sovereignty'!V5:V67,\"N/A\")", "High"),
    ]
    
    row += 1
    for risk, formula, severity in risks:
        ws.cell(row=row, column=1, value=risk)
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=3, value=severity)
        if severity == "Critical":
            ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif severity == "High":
            ws.cell(row=row, column=3).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A3"


# --- END BLOCK 3 ---
logger.info("Block 3 loaded: Individual Sheet Creators + Jurisdictional Risk + Dashboard (Sections 7-8)")

# ============================================================================
# SECTION 9: EVIDENCE REGISTER
# ============================================================================

def create_9_evidence_register(ws, styles):
    """Sheet 9: Evidence Register for vendor documentation."""
    
    # Header
    ws.merge_cells("A1:J1")
    ws["A1"] = "EVIDENCE REGISTER - VENDOR DUE DILIGENCE & CONTRACTS"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells("A2:J2")
    ws["A2"] = "Track all vendor contracts, certifications, and compliance documentation"
    apply_style(ws["A2"], styles["subheader"], "subheader")
    
    # Column headers
    headers = [
        ("Evidence ID", 15),
        ("Vendor Name", 25),
        ("Evidence Type", 25),
        ("Description", 40),
        ("File Location / URL", 40),
        ("Document Date", 14),
        ("Expiry/Renewal Date", 16),
        ("Document Owner", 20),
        ("Status", 14),
        ("Notes", 30),
    ]
    
    row = 4
    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        apply_style(cell, styles["column_header"], "header")
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    
    # Evidence Type dropdown
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Contract (MSA),Contract (DPA),Contract (SLA),ISO 27001 Cert,SOC 2 Report,FedRAMP Letter,Vendor Questionnaire,Jurisdictional Assessment,Risk Acceptance,Audit Report,Insurance Cert,Other"',
        allow_blank=True
    )
    ws.add_data_validation(ev_type_dv)
    
    # Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Current,Expired,Pending Renewal,Under Review,Superseded"',
        allow_blank=True
    )
    ws.add_data_validation(status_dv)
    
    # Pre-fill Evidence IDs and format input rows
    for i in range(1, 101):
        data_row = row + i
        # Auto-generated Evidence ID
        ws.cell(row=data_row, column=1, value=f"EV-VDD-{i:03d}")
        
        # Yellow input cells for remaining columns
        for col_idx in range(2, 11):
            cell = ws.cell(row=data_row, column=col_idx, value="")
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Apply dropdowns
    ev_type_dv.add("C5:C104")
    status_dv.add("I5:I104")
    
    # Evidence Checklist
    checklist_row = 108
    ws[f"A{checklist_row}"] = "EVIDENCE CHECKLIST"
    ws[f"A{checklist_row}"].font = Font(bold=True, size=11)
    
    checklist = [
        "☐ All MSAs uploaded to contract management system",
        "☐ DPAs reviewed by Legal and Data Protection Officer",
        "☐ SLAs attached to vendor records",
        "☐ Current ISO 27001/SOC 2 reports on file (<12 months old)",
        "☐ FedRAMP authorization letters archived (if applicable)",
        "☐ Vendor security questionnaires completed annually",
        "☐ Cyber insurance certificates verified (>$1M coverage)",
        "☐ Jurisdictional risk assessments documented for US-nexus providers",
        "☐ Risk acceptance forms signed for high-risk vendors",
    ]
    
    checklist_row += 1
    for item in checklist:
        ws[f"A{checklist_row}"] = item
        checklist_row += 1

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 10: APPROVAL SIGN-OFF (MULTI-STAKEHOLDER)
# ============================================================================

def create_10_approval_signoff(ws, styles):
    """Sheet 10: Multi-stakeholder Approval Sign-Off."""
    
    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "APPROVAL SIGN-OFF - VENDOR DUE DILIGENCE ASSESSMENT"
    apply_style(ws["A1"], styles["header"], "header")
    ws.row_dimensions[1].height = 35

    # Subtitle (row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = "Vendor Due Diligence & Contracts | Approval and sign-off for ISO 27001:2022 Control A.5.23"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10)
    ws["A2"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    row = 3
    
    # ASSESSMENT SUMMARY
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    
    completion_fields = [
        ("Assessment Completed By:", "", "input"),
        ("Completion Date:", "", "input"),
        ("Total Vendors Assessed:", "='8. Summary Dashboard'!B10", "formula"),
        ("Overall Compliance %:", "='8. Summary Dashboard'!G10", "formula"),
        ("Jurisdictional Risks Identified:", "='8. Summary Dashboard'!B20", "formula"),
        ("Critical Issues Identified:", "", "input"),
        ("Remediation Plan Attached:", "", "dropdown"),
    ]
    
    row += 1
    for label, value, field_type in completion_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        if field_type == "formula":
            ws[f"B{row}"] = value
        else:
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            ws[f"B{row}"].border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        row += 1

    # Dropdown for Remediation Plan
    dv_plan = DataValidation(type="list", formula1='"Yes,No,In Progress"', allow_blank=False)
    ws.add_data_validation(dv_plan)
    dv_plan.add(f"B{row-1}")
    
    # LEGAL REVIEW
    row += 2
    ws[f"A{row}"] = "LEGAL REVIEW"
    ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")
    
    row += 1
    thin = Side(style="thin")
    for label in ["Reviewed By (Legal Counsel):", "Review Date:", "Contract Compliance Status:", "Jurisdictional Risk Acceptance:", "Legal Comments:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1
    
    dv_legal = DataValidation(type="list", formula1='"Approved,Requires Negotiation,Non-Compliant"', allow_blank=False)
    ws.add_data_validation(dv_legal)
    dv_legal.add(f"B{row-3}")  # Contract Compliance Status
    
    dv_risk_accept = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected,N/A"', allow_blank=False)
    ws.add_data_validation(dv_risk_accept)
    dv_risk_accept.add(f"B{row-2}")  # Jurisdictional Risk Acceptance
    
    # PROCUREMENT REVIEW
    row += 2
    ws[f"A{row}"] = "PROCUREMENT REVIEW"
    ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")
    
    row += 1
    for label in ["Reviewed By (Procurement):", "Review Date:", "Commercial Terms Status:", "Procurement Comments:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1
    
    dv_proc = DataValidation(type="list", formula1='"Approved,Under Negotiation,Rejected"', allow_blank=False)
    ws.add_data_validation(dv_proc)
    dv_proc.add(f"B{row-2}")  # Commercial Terms Status
    
    # DATA PROTECTION OFFICER REVIEW
    row += 2
    ws[f"A{row}"] = "DATA PROTECTION OFFICER REVIEW"
    ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")
    
    row += 1
    for label in ["Reviewed By (DPO):", "Review Date:", "Data Protection Compliance:", "Cross-Border Transfer Status:", "DPO Comments:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

    dv_dpo = DataValidation(type="list", formula1='"Compliant,Partially Compliant,Non-Compliant"', allow_blank=False)
    ws.add_data_validation(dv_dpo)
    dv_dpo.add(f"B{row-3}")  # Data Protection Compliance
    
    dv_transfer = DataValidation(type="list", formula1='"Approved,Approved with SCCs,Requires TIA,Rejected"', allow_blank=False)
    ws.add_data_validation(dv_transfer)
    dv_transfer.add(f"B{row-2}")  # Cross-Border Transfer Status
    
    # CISO APPROVAL
    row += 2
    ws[f"A{row}"] = "CISO APPROVAL"
    ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")
    
    row += 1
    for label in ["Approved By (CISO):", "Approval Date:", "Security Risk Assessment:", "Jurisdictional Risk Acceptance:", "Executive Comments:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

    dv_ciso = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected"', allow_blank=False)
    ws.add_data_validation(dv_ciso)
    dv_ciso.add(f"B{row-3}")  # Security Risk Assessment
    dv_ciso.add(f"B{row-2}")  # Jurisdictional Risk Acceptance

    # Final decision
    row += 1
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

    decision_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(ws[f"B{row}"])

    # NEXT REVIEW DETAILS
    row += 3
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        thin = Side(style="thin")
        ws[f"B{row}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 11: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution - orchestrates workbook creation."""
    logger.info("=" * 70)
    logger.info("ISMS-IMP-A.5.23.S2 - Vendor Due Diligence & Contracts Generator")
    logger.info("Version 1.0 - WITH REGULATORY UPDATES (DORA/NIS2/AI Act/CLOUD Act)")
    logger.info("ISO/IEC 27001:2022 Control A.5.23: Cloud Services Security")
    logger.info("=" * 70)
    logger.info("\n'The first principle is that you must not fool yourself")
    logger.info(" - and you are the easiest person to fool.' - Richard Feynman")
    logger.info("\n(Translation: Evidence-based compliance prevents cloud washing!)\n")
    
    wb = create_workbook()
    styles = setup_styles()
    
    logger.info("[1/10] Creating Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"])
    
    logger.info("[2/10] Creating 2. Security Certifications sheet...")
    create_2_security_certifications(wb["2. Security Certifications"], styles)
    
    logger.info("[3/10] Creating 3. Contract Terms sheet (with DORA/NIS2)...")
    create_3_contract_terms(wb["3. Contract Terms"], styles)
    
    logger.info("[4/10] Creating 4. SLA Performance sheet...")
    create_4_sla_performance(wb["4. SLA Performance"], styles)
    
    logger.info("[5/10] Creating 5. Data Sovereignty sheet...")
    create_5_data_sovereignty(wb["5. Data Sovereignty"], styles)
    
    logger.info("[6/10] Creating 6. Forensics & Audit sheet...")
    create_6_forensics_audit(wb["6. Forensics & Audit"], styles)
    
    logger.info("[7/10] Creating 7. Jurisdictional Risk sheet...")
    create_7_jurisdictional_risk(wb["7. Jurisdictional Risk"], styles)
    
    logger.info("[8/10] Creating 8. Summary Dashboard sheet...")
    create_8_summary_dashboard(wb["8. Summary Dashboard"], styles)
    
    logger.info("[9/10] Creating 9. Evidence Register sheet...")
    create_9_evidence_register(wb["9. Evidence Register"], styles)
    
    logger.info("[10/10] Creating 10. Approval Sign-Off sheet...")
    create_10_approval_signoff(wb["10. Approval Sign-Of"], styles)
    
    # Save workbook
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"ISMS-IMP-A.5.23.S2_VendorDD_{timestamp}.xlsx"
    _wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
    _wkbk_dir.mkdir(exist_ok=True)
    wb.save(_wkbk_dir / filename)

    logger.info("\n" + "=" * 70)
    logger.info(f"{CHECK} Workbook generated successfully: {filename}")
    logger.info("=" * 70)
    logger.info("\nWorkbook contains 10 sheets:")
    logger.info("  1.  Instructions & Legend (with regulatory guidance)")
    logger.info("  2.  Security Certifications (ISO 27001, SOC 2, FedRAMP)")
    logger.info("  3.  Contract Terms (DPA, liability, DORA Art.30, NIS2)")
    logger.info("  4.  SLA Performance (availability, support, incidents)")
    logger.info("  5.  Data Sovereignty (GDPR, nFADP, SCCs, regions)")
    logger.info("  6.  Forensics & Audit (audit rights, IR, breach notification)")
    logger.info("  7.  Jurisdictional Risk")
    logger.info("  8.  Summary Dashboard (compliance metrics + jurisdictional KPIs)")
    logger.info("  9.  Evidence Register (contract & cert tracking)")
    logger.info("  10. Approval Sign-Off (Legal, Procurement, DPO, CISO)")
    logger.info("  • DORA Article 30 contract clause checklist")
    logger.info("  • NIS2 supply chain security requirements")
    logger.info("  • CLOUD Act jurisdictional risk assessment")
    logger.info("  • EU AI Act evaluation criteria")
    logger.info("  • DPO sign-off section")
    logger.info("\nNext steps:")
    logger.info(f"  1. Run validation: python3 excel_sanity_check_a523.py {filename}")
    logger.info("  2. Open in Excel and verify dropdowns work")
    logger.info("  3. Distribute to stakeholders for completion")
    logger.info("  4. Focus on Sheet 7 for US-nexus cloud providers")
    
    return filename


if __name__ == "__main__":
    main()


# --- END BLOCK 4 ---
logger.info("Block 4 loaded: Evidence Register, Approval Sign-Off, Main Execution (Sections 9-11)")
logger.info("\n" + "=" * 70)
logger.info("ALL BLOCKS LOADED - Script ready to execute!")
logger.info("=" * 70)
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
