#!/usr/bin/env python3
"""
ISMS-IMP-A.5.23 - Sample Data Population Script
==============================================
Populates all 4 cloud services assessment workbooks (S1-S4) with realistic
sample data for pipeline testing and PostgreSQL import validation.

Usage:
    python3 populate_sample_data_all.py

Pipeline position: Step 2 of 4
    1. Generate (S1-S5 generators)
    2. Fill in (this script) ← YOU ARE HERE
    3. Normalise (normalize_assessment_files_reg_a523.py)
    4. Consolidate (consolidate_reg_a523_dashboard.py)

Sample data represents a typical Swiss mid-sized company using:
- Microsoft 365, Azure, AWS for core cloud services
- Various SaaS tools (Salesforce, ServiceNow, etc.)
- Compliance focus: GDPR, FADP, DORA, NIS2
"""

import os
import glob
import logging
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =============================================================================
# LOGGING SETUP
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# CONSTANTS
# =============================================================================
STATUS_COMPLIANT = "✅ Compliant"
STATUS_PARTIAL = "⚠️ Partial"
STATUS_NON_COMPLIANT = "❌ Non-Compliant"

TODAY = datetime.now()
TODAY_STR = TODAY.strftime('%Y-%m-%d')

# =============================================================================
# SAMPLE DATA - CONSISTENT ACROSS ALL WORKBOOKS
# =============================================================================

SAMPLE_SERVICES = [
    {
        'name': 'Microsoft 365',
        'type': 'SaaS',
        'provider': 'Microsoft',
        'criticality': 'Critical',
        'data_class': 'Confidential',
        'region': 'Switzerland',
        'contract': 'Active',
        'exit_strategy': 'Cloud-to-Cloud',
        'alternative': 'Google Workspace',
        'export_format': 'PST/CSV',
        'export_tested': 'Yes',
        'lock_in': 'Medium',
        'dora_scope': 'Yes',
        'nis2_scope': 'Yes',
        'ai_risk': 'Limited Risk',
        'cloud_act': 'Potential Exposure (US HQ)',
        'hq_jurisdiction': 'United States',
        'eu_boundary': 'Yes',
        'cost_monthly': 45000,
        'cost_annual': 540000,
        'users': 500,
        'integrations': 12,
    },
    {
        'name': 'Azure Cloud Platform',
        'type': 'IaaS',
        'provider': 'Microsoft',
        'criticality': 'Critical',
        'data_class': 'Restricted',
        'region': 'Switzerland',
        'contract': 'Active',
        'exit_strategy': 'Hybrid',
        'alternative': 'AWS/GCP',
        'export_format': 'JSON/ARM Templates',
        'export_tested': 'Yes',
        'lock_in': 'High',
        'dora_scope': 'Yes',
        'nis2_scope': 'Yes',
        'ai_risk': 'Not Applicable',
        'cloud_act': 'Potential Exposure (US HQ)',
        'hq_jurisdiction': 'United States',
        'eu_boundary': 'Yes',
        'cost_monthly': 85000,
        'cost_annual': 1020000,
        'users': 200,
        'integrations': 25,
    },
    {
        'name': 'AWS S3 Storage',
        'type': 'Storage',
        'provider': 'Amazon',
        'criticality': 'High',
        'data_class': 'Confidential',
        'region': 'EU/EEA',
        'contract': 'Active',
        'exit_strategy': 'Cloud-to-Cloud',
        'alternative': 'Azure Blob',
        'export_format': 'Native/S3 API',
        'export_tested': 'Yes',
        'lock_in': 'Low',
        'dora_scope': 'Yes',
        'nis2_scope': 'Yes',
        'ai_risk': 'Not Applicable',
        'cloud_act': 'Potential Exposure (US HQ)',
        'hq_jurisdiction': 'United States',
        'eu_boundary': 'No',
        'cost_monthly': 12000,
        'cost_annual': 144000,
        'users': 50,
        'integrations': 8,
    },
    {
        'name': 'Salesforce CRM',
        'type': 'SaaS',
        'provider': 'Salesforce',
        'criticality': 'High',
        'data_class': 'Confidential',
        'region': 'EU/EEA',
        'contract': 'Active',
        'exit_strategy': 'Cloud-to-Cloud',
        'alternative': 'Microsoft Dynamics',
        'export_format': 'CSV/Data Loader',
        'export_tested': 'Yes',
        'lock_in': 'Medium',
        'dora_scope': 'No',
        'nis2_scope': 'No',
        'ai_risk': 'Limited Risk',
        'cloud_act': 'Potential Exposure (US HQ)',
        'hq_jurisdiction': 'United States',
        'eu_boundary': 'No',
        'cost_monthly': 18000,
        'cost_annual': 216000,
        'users': 120,
        'integrations': 6,
    },
    {
        'name': 'ServiceNow ITSM',
        'type': 'SaaS',
        'provider': 'ServiceNow',
        'criticality': 'High',
        'data_class': 'Internal',
        'region': 'EU/EEA',
        'contract': 'Active',
        'exit_strategy': 'Cloud-to-Cloud',
        'alternative': 'Jira Service Mgmt',
        'export_format': 'XML/CSV Export',
        'export_tested': 'Yes',
        'lock_in': 'Medium',
        'dora_scope': 'Yes',
        'nis2_scope': 'Yes',
        'ai_risk': 'Not Applicable',
        'cloud_act': 'Potential Exposure (US HQ)',
        'hq_jurisdiction': 'United States',
        'eu_boundary': 'No',
        'cost_monthly': 15000,
        'cost_annual': 180000,
        'users': 300,
        'integrations': 10,
    },
    {
        'name': 'SAP S/4HANA Cloud',
        'type': 'SaaS',
        'provider': 'SAP',
        'criticality': 'Critical',
        'data_class': 'Restricted',
        'region': 'Switzerland',
        'contract': 'Active',
        'exit_strategy': 'Hybrid',
        'alternative': 'None viable',
        'export_format': 'RFC/IDoc',
        'export_tested': 'Partial',
        'lock_in': 'Critical',
        'dora_scope': 'Yes',
        'nis2_scope': 'Yes',
        'ai_risk': 'Not Applicable',
        'cloud_act': 'No Exposure',
        'hq_jurisdiction': 'EU/EEA',
        'eu_boundary': 'Yes',
        'cost_monthly': 95000,
        'cost_annual': 1140000,
        'users': 400,
        'integrations': 18,
    },
    {
        'name': 'Crowdstrike Falcon',
        'type': 'Security',
        'provider': 'Crowdstrike',
        'criticality': 'Critical',
        'data_class': 'Confidential',
        'region': 'EU/EEA',
        'contract': 'Active',
        'exit_strategy': 'Cloud-to-Cloud',
        'alternative': 'Microsoft Defender',
        'export_format': 'JSON/SIEM',
        'export_tested': 'Yes',
        'lock_in': 'Low',
        'dora_scope': 'Yes',
        'nis2_scope': 'Yes',
        'ai_risk': 'Limited Risk',
        'cloud_act': 'Potential Exposure (US HQ)',
        'hq_jurisdiction': 'United States',
        'eu_boundary': 'No',
        'cost_monthly': 8000,
        'cost_annual': 96000,
        'users': 500,
        'integrations': 5,
    },
    {
        'name': 'Slack Enterprise',
        'type': 'SaaS',
        'provider': 'Salesforce',
        'criticality': 'Medium',
        'data_class': 'Internal',
        'region': 'EU/EEA',
        'contract': 'Active',
        'exit_strategy': 'Cloud-to-Cloud',
        'alternative': 'Microsoft Teams',
        'export_format': 'JSON Export',
        'export_tested': 'Yes',
        'lock_in': 'Low',
        'dora_scope': 'No',
        'nis2_scope': 'No',
        'ai_risk': 'Not Applicable',
        'cloud_act': 'Potential Exposure (US HQ)',
        'hq_jurisdiction': 'United States',
        'eu_boundary': 'No',
        'cost_monthly': 5000,
        'cost_annual': 60000,
        'users': 500,
        'integrations': 8,
    },
]

VENDOR_CERTS = {
    'Microsoft': {'iso27001': True, 'iso_num': 'IS 745336', 'soc2': True, 'fedramp': 'Yes (High)', 'other': 'C5, CSA STAR'},
    'Amazon': {'iso27001': True, 'iso_num': 'IS 624562', 'soc2': True, 'fedramp': 'Yes (Moderate)', 'other': 'C5, CSA STAR'},
    'Salesforce': {'iso27001': True, 'iso_num': 'IS 589012', 'soc2': True, 'fedramp': 'No', 'other': 'C5'},
    'ServiceNow': {'iso27001': True, 'iso_num': 'IS 612988', 'soc2': True, 'fedramp': 'No', 'other': 'CSA STAR'},
    'SAP': {'iso27001': True, 'iso_num': 'IS 701245', 'soc2': True, 'fedramp': 'No', 'other': 'C5, BSI'},
    'Crowdstrike': {'iso27001': True, 'iso_num': 'IS 698234', 'soc2': True, 'fedramp': 'Yes (Moderate)', 'other': 'SOC 3'},
}


def safely_write(ws, row, col, value):
    """Write value to cell, skipping merged cells."""
    try:
        ws.cell(row=row, column=col, value=value)
    except Exception as e:
        pass


def write_row(ws, row, data_dict):
    """Write a dict of {col_letter: value} to a row."""
    for col_letter, value in data_dict.items():
        col_num = ord(col_letter[0]) - ord('A') + 1
        if len(col_letter) > 1:
            col_num = (ord(col_letter[0]) - ord('A') + 1) * 26 + (ord(col_letter[1]) - ord('A') + 1)
        safely_write(ws, row, col_num, value)


# =============================================================================
# S1 - INVENTORY WORKBOOK POPULATION
# =============================================================================

def populate_s1_inventory(wb):
    """Populate S1 Cloud Service Inventory workbook.

    Sheet structure (from generator):
      Sheets 2/3/5: DATA_START_ROW=8, cols A-AG (33 cols)
      Sheet 4 (Exit): DATA_START_ROW=4, cols A-AC (29 cols)
      Sheet 6 (Data Class): DATA_START_ROW=4, cols A-J (10 cols)
      Sheet 7 (Criticality): DATA_START_ROW=4, cols A-L (12 cols)
      Sheet 9 (Evidence): DATA_START_ROW=5, cols A-I (9 cols)
    """
    logger.info("Populating S1 - Cloud Service Inventory...")

    # --- Sheet 2: SaaS Services (row 8, cols A-AG) ---
    ws = wb['2. SaaS Services']
    saas = [s for s in SAMPLE_SERVICES if s['type'] == 'SaaS']
    for i, svc in enumerate(saas):
        r = 8 + i
        ws.cell(row=r, column=1, value=svc['name'])          # A: Cloud_Service_Name
        ws.cell(row=r, column=2, value=svc['type'])           # B: Service_Type
        ws.cell(row=r, column=3, value=svc['provider'])       # C: Vendor_Name
        ws.cell(row=r, column=4, value=svc['criticality'])    # D: Service_Criticality
        ws.cell(row=r, column=5, value=svc['data_class'])     # E: Data_Classification
        ws.cell(row=r, column=6, value=svc['region'])         # F: Data_Residency
        ws.cell(row=r, column=7, value=svc['contract'])       # G: Contract_Status
        ws.cell(row=r, column=8, value=STATUS_COMPLIANT)      # H: Status
        ws.cell(row=r, column=9, value='SharePoint Evidence Library')  # I: Evidence_Location
        ws.cell(row=r, column=15, value='IT Operations')      # O: Service_Owner
        ws.cell(row=r, column=18, value=svc['cost_monthly'])  # R: Monthly_Cost_CHF
        ws.cell(row=r, column=19, value=svc['cost_annual'])   # S: Annual_Contract_Value
        ws.cell(row=r, column=20, value=svc['users'])         # T: Licensed_Users
        ws.cell(row=r, column=21, value=svc['integrations'])  # U: Integration_Count
        ws.cell(row=r, column=22, value='Yes')                # V: Backup_Available
        ws.cell(row=r, column=23, value='Easy (<=30 days)')   # W: Exit_Feasibility
        ws.cell(row=r, column=24, value=TODAY_STR)            # X: Last_Review_Date
        ws.cell(row=r, column=25, value=svc['hq_jurisdiction'])  # Y: Provider_HQ_Jurisdiction
        ws.cell(row=r, column=26, value=svc['cloud_act'])     # Z: CLOUD_Act_Exposure
        ws.cell(row=r, column=27, value=svc['region'])        # AA: Data_Processing_Locations
        ws.cell(row=r, column=28, value=svc['eu_boundary'])   # AB: EU_Data_Boundary
        ws.cell(row=r, column=29, value=svc['dora_scope'])    # AC: DORA_In_Scope
        ws.cell(row=r, column=30, value=svc['nis2_scope'])    # AD: NIS2_In_Scope
        ws.cell(row=r, column=31, value=svc['ai_risk'])       # AE: AI_Classification
        ws.cell(row=r, column=32, value=svc['lock_in'] + ' (Limited alternatives)' if svc['lock_in'] == 'Medium' else 'Low (Multiple alternatives)')  # AF
        ws.cell(row=r, column=33, value='Yes')                # AG: Alternatives_Identified
    logger.info(f"   [OK] SaaS Services: {len(saas)} entries")

    # --- Sheet 3: IaaS/PaaS Services (row 8, cols A-AG) ---
    ws = wb['3. IaaS PaaS Services']
    iaas = [s for s in SAMPLE_SERVICES if s['type'] in ['IaaS', 'PaaS']]
    for i, svc in enumerate(iaas):
        r = 8 + i
        ws.cell(row=r, column=1, value=svc['name'])
        ws.cell(row=r, column=2, value=svc['type'])
        ws.cell(row=r, column=3, value=svc['provider'])
        ws.cell(row=r, column=4, value=svc['criticality'])
        ws.cell(row=r, column=5, value=svc['data_class'])
        ws.cell(row=r, column=6, value=svc['region'])
        ws.cell(row=r, column=7, value=svc['contract'])
        ws.cell(row=r, column=8, value=STATUS_COMPLIANT)
        ws.cell(row=r, column=9, value='SharePoint Evidence Library')
        ws.cell(row=r, column=15, value='Cloud Engineering')
        ws.cell(row=r, column=18, value=svc['cost_monthly'])
        ws.cell(row=r, column=19, value=svc['cost_annual'])
        ws.cell(row=r, column=20, value=svc['users'])
        ws.cell(row=r, column=21, value=svc['integrations'])
        ws.cell(row=r, column=22, value='Yes')
        ws.cell(row=r, column=23, value='Medium (31-90 days)')
        ws.cell(row=r, column=24, value=TODAY_STR)
        ws.cell(row=r, column=25, value=svc['hq_jurisdiction'])
        ws.cell(row=r, column=26, value=svc['cloud_act'])
        ws.cell(row=r, column=29, value=svc['dora_scope'])
        ws.cell(row=r, column=30, value=svc['nis2_scope'])
    logger.info(f"   [OK] IaaS/PaaS Services: {len(iaas)} entries")

    # --- Sheet 4: Exit Strategy (row 4, cols A-AC, 29 cols) ---
    ws = wb['4. Cloud Security Services']
    for i, svc in enumerate(SAMPLE_SERVICES):
        r = 4 + i
        ws.cell(row=r, column=1, value=svc['name'])           # A: Service Name
        ws.cell(row=r, column=2, value=svc['provider'])        # B: Provider
        ws.cell(row=r, column=3, value=svc['type'])            # C: Service Type
        ws.cell(row=r, column=4, value='Production')           # D: Environment
        ws.cell(row=r, column=5, value='J. Mueller')           # E: Business Owner
        ws.cell(row=r, column=6, value='T. Schmidt')           # F: Technical Owner
        ws.cell(row=r, column=7, value=svc['data_class'])      # G: Data Classification
        ws.cell(row=r, column=8, value=svc['criticality'])     # H: Criticality
        ws.cell(row=r, column=9, value=svc['users'])           # I: Users
        ws.cell(row=r, column=10, value=svc['cost_annual'])    # J: Cost (Annual CHF)
        ws.cell(row=r, column=11, value='2027-01-14')          # K: Contract End Date
        ws.cell(row=r, column=12, value=svc['region'])         # L: Primary Region
        ws.cell(row=r, column=13, value='EU/EEA')              # M: Backup Region
        ws.cell(row=r, column=14, value=svc['integrations'])   # N: Integration Count
        ws.cell(row=r, column=15, value='Yes')                 # O: API Dependency
        ws.cell(row=r, column=16, value='ISO 27001, SOC 2')    # P: Compliance Certs
        ws.cell(row=r, column=17, value=STATUS_COMPLIANT)      # Q: Current Status
        ws.cell(row=r, column=18, value=svc['exit_strategy'])  # R: Exit Strategy Type
        ws.cell(row=r, column=19, value=svc['alternative'])    # S: Alternative Identified
        ws.cell(row=r, column=20, value=svc['export_format'])  # T: Export Format Available
        ws.cell(row=r, column=21, value=svc['export_tested'])  # U: Export Tested
        ws.cell(row=r, column=22, value=500)                   # V: Data Volume (GB)
        complexity = 'Cloud-to-Cloud (Low)' if svc['lock_in'] == 'Low' else ('Cloud-to-Cloud (Medium)' if svc['lock_in'] == 'Medium' else 'Hybrid (High)')
        ws.cell(row=r, column=23, value=complexity)            # W: Migration Complexity
        ws.cell(row=r, column=24, value=svc['lock_in'])        # X: Lock-In Risk
    logger.info(f"   [OK] Exit Strategy: {len(SAMPLE_SERVICES)} entries")

    # --- Sheet 5: Cloud Storage (row 8, cols A-AG) ---
    ws = wb['5. Cloud Storage Services']
    storage = [s for s in SAMPLE_SERVICES if s['type'] == 'Storage']
    for i, svc in enumerate(storage):
        r = 8 + i
        ws.cell(row=r, column=1, value=svc['name'])
        ws.cell(row=r, column=2, value=svc['type'])
        ws.cell(row=r, column=3, value=svc['provider'])
        ws.cell(row=r, column=4, value=svc['criticality'])
        ws.cell(row=r, column=5, value=svc['data_class'])
        ws.cell(row=r, column=6, value=svc['region'])
        ws.cell(row=r, column=7, value=svc['contract'])
        ws.cell(row=r, column=8, value=STATUS_COMPLIANT)
        ws.cell(row=r, column=9, value='SharePoint Evidence Library')
        ws.cell(row=r, column=15, value='Cloud Engineering')
        ws.cell(row=r, column=18, value=svc['cost_monthly'])
        ws.cell(row=r, column=19, value=svc['cost_annual'])
    logger.info(f"   [OK] Cloud Storage: {len(storage)} entries")

    # --- Sheet 6: Data Classification (row 4, cols A-J) ---
    ws = wb['6. Data Classification']
    data_categories = [
        ('Customer PII', 'Personal data of customers', 'Restricted', 'Microsoft 365', 'SharePoint', 'Legal', '7 years', 'Yes - Personal', 'EEA Only'),
        ('Employee HR Data', 'Employee records and payroll', 'Restricted', 'SAP S/4HANA', 'SAP Cloud', 'HR Director', '10 years', 'Yes - Special Category', 'None'),
        ('Financial Reports', 'Quarterly and annual financials', 'Confidential', 'SAP S/4HANA', 'SAP Cloud', 'CFO', '10 years', 'No', 'None'),
        ('Sales Data', 'CRM and pipeline data', 'Confidential', 'Salesforce CRM', 'Salesforce', 'Sales Director', '5 years', 'Yes - Personal', 'Adequacy Decision'),
        ('IT Logs', 'System and security logs', 'Internal', 'Azure Cloud', 'Azure Log Analytics', 'CISO', '1 year', 'No', 'EEA Only'),
    ]
    for i, cat in enumerate(data_categories):
        r = 4 + i
        ws.cell(row=r, column=1, value=f'DC-{i+1:03d}')  # A: Data_ID
        ws.cell(row=r, column=2, value=cat[0])             # B: Data Category
        ws.cell(row=r, column=3, value=cat[1])             # C: Description
        ws.cell(row=r, column=4, value=cat[2])             # D: Classification Level
        ws.cell(row=r, column=5, value=cat[3])             # E: Primary Cloud Service
        ws.cell(row=r, column=6, value=cat[4])             # F: Storage Location
        ws.cell(row=r, column=7, value=cat[5])             # G: Data Owner
        ws.cell(row=r, column=8, value=cat[6])             # H: Retention Period
        ws.cell(row=r, column=9, value=cat[7])             # I: Personal Data (GDPR)
        ws.cell(row=r, column=10, value=cat[8])            # J: Cross-Border Transfer
    logger.info(f"   [OK] Data Classification: {len(data_categories)} entries")

    # --- Sheet 7: Service Criticality (row 4, cols A-L) ---
    ws = wb['7. Service Criticality']
    for i, svc in enumerate(SAMPLE_SERVICES):
        r = 4 + i
        rto = 4 if svc['criticality'] == 'Critical' else (8 if svc['criticality'] == 'High' else 24)
        rpo = 1 if svc['criticality'] == 'Critical' else (4 if svc['criticality'] == 'High' else 24)
        ws.cell(row=r, column=1, value=f'SC-{i+1:03d}')                     # A: Service_ID
        ws.cell(row=r, column=2, value=svc['name'])                          # B: Cloud Service Name
        ws.cell(row=r, column=3, value=svc['type'])                          # C: Service Category
        ws.cell(row=r, column=4, value='Core Business Operations')           # D: Business Process
        ws.cell(row=r, column=5, value=svc['criticality'])                   # E: Criticality Level
        ws.cell(row=r, column=6, value=rto)                                  # F: RTO (Hours)
        ws.cell(row=r, column=7, value=rpo)                                  # G: RPO (Hours)
        ws.cell(row=r, column=8, value=rto * 3)                              # H: MTPD (Hours)
        ws.cell(row=r, column=9, value='No' if svc['alternative'] != 'None viable' else 'Yes')  # I
        ws.cell(row=r, column=10, value='Yes' if svc['alternative'] != 'None viable' else 'No')  # J
        ws.cell(row=r, column=11, value='In-Scope' if svc['dora_scope'] == 'Yes' else 'Out-of-Scope')  # K
        ws.cell(row=r, column=12, value='P1 - Critical' if svc['criticality'] == 'Critical' else 'P2 - High')  # L
    logger.info(f"   [OK] Service Criticality: {len(SAMPLE_SERVICES)} entries")

    # --- Sheet 9: Evidence Register (row 5, cols A-I) ---
    ws = wb['9. Evidence Register']
    evidence_items = [
        ('EV-INV-001', 'Microsoft 365', 'CMDB Export', 'Full SaaS inventory export from ServiceNow CMDB', '/evidence/s1/cmdb_export_2025.xlsx', 'IT Operations', '1 year', 'Current'),
        ('EV-INV-002', 'Azure Cloud Platform', 'Config File', 'Azure resource group inventory JSON export', '/evidence/s1/azure_resources.json', 'Cloud Engineering', '1 year', 'Current'),
        ('EV-INV-003', 'AWS S3 Storage', 'Screenshot', 'AWS console S3 bucket listing with encryption status', '/evidence/s1/aws_s3_screenshot.png', 'Cloud Engineering', '6 months', 'Current'),
        ('EV-INV-004', 'Salesforce CRM', 'Contract', 'Salesforce Enterprise Agreement with DPA', '/evidence/s1/sf_contract_2024.pdf', 'Procurement', '3 years', 'Current'),
        ('EV-INV-005', 'SAP S/4HANA Cloud', 'Risk Assessment', 'SAP cloud risk assessment and exit feasibility', '/evidence/s1/sap_risk_assessment.pdf', 'Risk Management', '1 year', 'Current'),
    ]
    for i, ev in enumerate(evidence_items):
        r = 5 + i
        ws.cell(row=r, column=1, value=ev[0])  # A: Evidence ID
        ws.cell(row=r, column=2, value=ev[1])  # B: Cloud Service
        ws.cell(row=r, column=3, value=ev[2])  # C: Evidence Type
        ws.cell(row=r, column=4, value=ev[3])  # D: Description
        ws.cell(row=r, column=5, value=ev[4])  # E: File Location
        ws.cell(row=r, column=6, value=TODAY_STR)  # F: Collection Date
        ws.cell(row=r, column=7, value=ev[5])  # G: Collected By
        ws.cell(row=r, column=8, value=ev[6])  # H: Retention
        ws.cell(row=r, column=9, value=ev[7])  # I: Status
    logger.info(f"   [OK] Evidence Register: {len(evidence_items)} entries")

    return wb


# =============================================================================
# S2 - VENDOR DUE DILIGENCE POPULATION
# =============================================================================

def populate_s2_vendor_dd(wb):
    """Populate S2 Vendor Due Diligence workbook.

    Sheet structure (from generator):
      Sheets 2-6: DATA_START_ROW=5, base cols A-Q (17) + extended R+
      Sheet 7 (Jurisdictional): DATA_START_ROW=7, cols A-T (20)
      Sheet 9 (Evidence): DATA_START_ROW=5, cols A-J (10)
    """
    logger.info("Populating S2 - Vendor Due Diligence...")

    # Unique vendors from services
    vendors = {}
    for svc in SAMPLE_SERVICES:
        if svc['provider'] not in vendors:
            vendors[svc['provider']] = svc

    # --- Sheet 2: Security Certifications (row 5, base A-Q + ext R-X) ---
    ws = wb['2. Security Certifications']
    for i, (vendor_name, svc) in enumerate(vendors.items()):
        r = 5 + i
        certs = VENDOR_CERTS.get(vendor_name, {})
        # Base columns A-Q
        ws.cell(row=r, column=1, value=svc['name'])           # A: Cloud Service Name
        ws.cell(row=r, column=2, value=vendor_name)            # B: Vendor Name
        ws.cell(row=r, column=3, value=svc['type'])            # C: Service Type
        ws.cell(row=r, column=4, value=svc['criticality'])     # D: Service Criticality
        ws.cell(row=r, column=5, value=svc['data_class'])      # E: Data Classification
        ws.cell(row=r, column=6, value='MSA + DPA')            # F: Contract Type
        ws.cell(row=r, column=7, value='2024-01-15')           # G: Contract Start Date
        ws.cell(row=r, column=8, value=STATUS_COMPLIANT)       # H: Status
        ws.cell(row=r, column=9, value='/evidence/s2/certs/')  # I: Evidence Location
        ws.cell(row=r, column=15, value=f'{vendor_name} Legal')  # O: Vendor Contact
        ws.cell(row=r, column=17, value='Procurement')         # Q: Contract Owner
        # Extended columns R-X
        ws.cell(row=r, column=18, value='Yes (Current)' if certs.get('iso27001') else 'No')  # R: ISO 27001
        ws.cell(row=r, column=19, value=certs.get('iso_num', ''))  # S: ISO Cert Number
        ws.cell(row=r, column=20, value='2026-06-15')          # T: ISO Expiry Date
        ws.cell(row=r, column=21, value='Yes (< 6 months)' if certs.get('soc2') else 'No')  # U: SOC 2
        ws.cell(row=r, column=22, value='2025-09-30')          # V: SOC 2 Report Date
        ws.cell(row=r, column=23, value=certs.get('fedramp', 'No'))  # W: FedRAMP
        ws.cell(row=r, column=24, value=certs.get('other', ''))  # X: Other Certs
    logger.info(f"   [OK] Security Certifications: {len(vendors)} vendors")

    # --- Sheet 3: Contract Terms (row 5, base A-Q + ext R-Z) ---
    ws = wb['3. Contract Terms']
    for i, svc in enumerate(SAMPLE_SERVICES):
        r = 5 + i
        ws.cell(row=r, column=1, value=svc['name'])           # A
        ws.cell(row=r, column=2, value=svc['provider'])        # B
        ws.cell(row=r, column=3, value=svc['type'])            # C
        ws.cell(row=r, column=4, value=svc['criticality'])     # D
        ws.cell(row=r, column=5, value=svc['data_class'])      # E
        ws.cell(row=r, column=6, value='MSA + DPA')            # F
        ws.cell(row=r, column=7, value='2024-01-15')           # G
        ws.cell(row=r, column=8, value=STATUS_COMPLIANT)       # H
        ws.cell(row=r, column=9, value='/evidence/s2/contracts/')  # I
        ws.cell(row=r, column=17, value='Legal')               # Q
        # Extended R-Z
        ws.cell(row=r, column=18, value='Yes (Adequate)')      # R: Data Protection Clause
        ws.cell(row=r, column=19, value='Yes (List Provided)') # S: Subprocessor Disclosure
        ws.cell(row=r, column=20, value='CHF 5,000,000')       # T: Liability Cap
        ws.cell(row=r, column=21, value='Yes (Favorable)')     # U: Indemnification
        ws.cell(row=r, column=22, value='<=30 days')           # V: Termination Notice
        ws.cell(row=r, column=23, value='Yes (30 days)')       # W: Data Return
        ws.cell(row=r, column=24, value='Yes (Opt-Out)')       # X: Auto-Renewal
        dora = 'Yes (Full)' if svc['dora_scope'] == 'Yes' else 'N/A (Not in scope)'
        ws.cell(row=r, column=25, value=dora)                  # Y: DORA Art.30
        ws.cell(row=r, column=26, value='Yes' if svc['nis2_scope'] == 'Yes' else 'N/A')  # Z: NIS2
    logger.info(f"   [OK] Contract Terms: {len(SAMPLE_SERVICES)} contracts")

    # --- Sheet 4: SLA Performance (row 5, base A-Q + ext R-X) ---
    ws = wb['4. SLA Performance']
    for i, svc in enumerate(SAMPLE_SERVICES):
        r = 5 + i
        ws.cell(row=r, column=1, value=svc['name'])
        ws.cell(row=r, column=2, value=svc['provider'])
        ws.cell(row=r, column=3, value=svc['type'])
        ws.cell(row=r, column=4, value=svc['criticality'])
        ws.cell(row=r, column=5, value=svc['data_class'])
        ws.cell(row=r, column=6, value='MSA + DPA')
        ws.cell(row=r, column=7, value='2024-01-15')
        ws.cell(row=r, column=8, value=STATUS_COMPLIANT)
        ws.cell(row=r, column=9, value='/evidence/s2/sla/')
        ws.cell(row=r, column=17, value='IT Operations')
        # Extended R-X
        ws.cell(row=r, column=18, value='99.95%')             # R: SLA Availability %
        ws.cell(row=r, column=19, value='99.97%')             # S: Actual Availability Q4
        ws.cell(row=r, column=20, value='None')               # T: SLA Credits Claimed
        ws.cell(row=r, column=21, value='P1: 15min, P2: 1hr') # U: Support Response Time
        ws.cell(row=r, column=22, value=2)                     # V: Incident Count Q4
        ws.cell(row=r, column=23, value='2.5 hours')           # W: Mean Time to Resolve
        ws.cell(row=r, column=24, value=TODAY_STR)             # X: SLA Review Date
    logger.info(f"   [OK] SLA Performance: {len(SAMPLE_SERVICES)} entries")

    # --- Sheet 5: Data Sovereignty (row 5, base A-Q + ext R-X) ---
    ws = wb['5. Data Sovereignty']
    for i, svc in enumerate(SAMPLE_SERVICES):
        r = 5 + i
        ws.cell(row=r, column=1, value=svc['name'])
        ws.cell(row=r, column=2, value=svc['provider'])
        ws.cell(row=r, column=3, value=svc['type'])
        ws.cell(row=r, column=4, value=svc['criticality'])
        ws.cell(row=r, column=5, value=svc['data_class'])
        ws.cell(row=r, column=6, value='MSA + DPA')
        ws.cell(row=r, column=7, value='2024-01-15')
        ws.cell(row=r, column=8, value=STATUS_COMPLIANT)
        ws.cell(row=r, column=9, value='/evidence/s2/sovereignty/')
        ws.cell(row=r, column=17, value='DPO')
        # Extended R-X
        ws.cell(row=r, column=18, value=svc['region'])        # R: Data Processing Region
        ch = 'Yes (Contractual)' if 'Switzerland' in svc['region'] else 'Yes (Technical)'
        ws.cell(row=r, column=19, value=ch)                    # S: Data Residency Verified
        scc = 'N/A' if 'Switzerland' in svc['region'] else 'Yes (EU SCC)'
        ws.cell(row=r, column=20, value=scc)                   # T: SCCs
        ws.cell(row=r, column=21, value='Adequacy Decision')   # U: Transfer Mechanism
        ws.cell(row=r, column=22, value='Yes (Certified)')     # V: GDPR Compliance
        ws.cell(row=r, column=23, value='Yes (Certified)')     # W: Swiss nFADP
        ws.cell(row=r, column=24, value='GDPR + nFADP')        # X: Regulatory Framework
    logger.info(f"   [OK] Data Sovereignty: {len(SAMPLE_SERVICES)} entries")

    # --- Sheet 6: Forensics & Audit (row 5, base A-Q + ext R-X) ---
    ws = wb['6. Forensics & Audit']
    for i, svc in enumerate(SAMPLE_SERVICES):
        r = 5 + i
        ws.cell(row=r, column=1, value=svc['name'])
        ws.cell(row=r, column=2, value=svc['provider'])
        ws.cell(row=r, column=3, value=svc['type'])
        ws.cell(row=r, column=4, value=svc['criticality'])
        ws.cell(row=r, column=5, value=svc['data_class'])
        ws.cell(row=r, column=6, value='MSA + DPA')
        ws.cell(row=r, column=7, value='2024-01-15')
        ws.cell(row=r, column=8, value=STATUS_COMPLIANT)
        ws.cell(row=r, column=9, value='/evidence/s2/forensics/')
        ws.cell(row=r, column=17, value='Security')
        # Extended R-X
        ws.cell(row=r, column=18, value='Yes (Full)')          # R: Forensics Support
        ws.cell(row=r, column=19, value='4 hours')             # S: Forensics SLA
        ws.cell(row=r, column=20, value='Yes (With Notice)')   # T: Right to Audit
        ws.cell(row=r, column=21, value='<=7 days')            # U: Audit Notice Period
        ws.cell(row=r, column=22, value='1-4 hours')           # V: Incident Notification SLA
        ws.cell(row=r, column=23, value='Yes (72 hours)')      # W: Breach Notification
        ws.cell(row=r, column=24, value='Yes')                 # X: IR Playbook Provided
    logger.info(f"   [OK] Forensics & Audit: {len(SAMPLE_SERVICES)} entries")

    # --- Sheet 7: Jurisdictional Risk (row 7, cols A-T, 20 cols) ---
    ws = wb['7. Jurisdictional Risk']
    for i, svc in enumerate(SAMPLE_SERVICES):
        r = 7 + i
        us_hq = svc['hq_jurisdiction'] == 'United States'
        ws.cell(row=r, column=1, value=f'JRA-{i+1:03d}')       # A: Assessment ID
        ws.cell(row=r, column=2, value=svc['name'])              # B: Cloud Service Name
        ws.cell(row=r, column=3, value=svc['provider'])          # C: Provider Name
        ws.cell(row=r, column=4, value='United States' if us_hq else 'Germany')  # D: Provider HQ Country
        ws.cell(row=r, column=5, value=svc['hq_jurisdiction'])   # E: Provider HQ Jurisdiction
        ws.cell(row=r, column=6, value='Yes' if us_hq else 'No')  # F: US Parent Company
        ws.cell(row=r, column=7, value=svc['cloud_act'])         # G: CLOUD Act Applicability
        ws.cell(row=r, column=8, value=svc['region'])            # H: Data Processing Locations
        ws.cell(row=r, column=9, value=svc['eu_boundary'])       # I: EU Data Boundary
        ws.cell(row=r, column=10, value='Yes' if svc['data_class'] == 'Restricted' else 'No')  # J: CMK
        ws.cell(row=r, column=11, value='Yes' if us_hq else 'No')  # K: Legal Challenge
        ws.cell(row=r, column=12, value='Adequacy Decision')     # L: Adequacy Status
        ws.cell(row=r, column=13, value='SCCs')                  # M: Transfer Mechanism
        risk = 'High' if us_hq and svc['data_class'] == 'Restricted' else ('Medium' if us_hq else 'Low')
        ws.cell(row=r, column=14, value=risk)                    # N: Risk Level
        ws.cell(row=r, column=15, value='CISO')                  # O: Risk Accepted By
        ws.cell(row=r, column=16, value=TODAY_STR)               # P: Risk Acceptance Date
        ws.cell(row=r, column=17, value='EU Data Boundary + CMK' if us_hq else 'N/A')  # Q
        ws.cell(row=r, column=18, value=TODAY_STR)               # R: Review Date
        ws.cell(row=r, column=19, value=f'EV-VDD-{i+1:03d}')    # S: Evidence Reference
        ws.cell(row=r, column=20, value='Annual review completed')  # T: Notes
    logger.info(f"   [OK] Jurisdictional Risk: {len(SAMPLE_SERVICES)} entries")

    # --- Sheet 9: Evidence Register (row 5, cols A-J, 10 cols) ---
    ws = wb['9. Evidence Register']
    evidence = [
        ('EV-VDD-001', 'Microsoft', 'ISO 27001 Cert', 'Microsoft ISO 27001:2022 certificate', '/evidence/s2/ms_iso27001.pdf', '2025-01-15', '2026-06-15', 'Compliance', 'Current', ''),
        ('EV-VDD-002', 'Microsoft', 'SOC 2 Report', 'SOC 2 Type II report for Azure', '/evidence/s2/ms_soc2_azure.pdf', '2025-03-01', '2026-03-01', 'Compliance', 'Current', ''),
        ('EV-VDD-003', 'Amazon', 'ISO 27001 Cert', 'AWS ISO 27001:2022 certificate', '/evidence/s2/aws_iso27001.pdf', '2025-02-01', '2026-07-01', 'Compliance', 'Current', ''),
        ('EV-VDD-004', 'Salesforce', 'Contract (DPA)', 'Salesforce Data Processing Agreement', '/evidence/s2/sf_dpa.pdf', '2024-06-15', '2027-01-14', 'Legal', 'Current', ''),
        ('EV-VDD-005', 'SAP', 'Vendor Questionnaire', 'SAP security questionnaire response', '/evidence/s2/sap_vsaq.pdf', '2025-01-10', '2026-01-10', 'Security', 'Current', ''),
    ]
    for i, ev in enumerate(evidence):
        r = 5 + i
        for j, val in enumerate(ev):
            ws.cell(row=r, column=j + 1, value=val)
    logger.info(f"   [OK] Evidence Register: {len(evidence)} entries")

    return wb


# =============================================================================
# S3 - SECURE CONFIGURATION POPULATION
# =============================================================================

def populate_s3_secure_config(wb):
    """Populate S3 Secure Configuration workbook.

    Sheet structure (from generator):
      Sheets 2-6: DATA_START_ROW=5, base cols A-Q (17) + extended R+
      Sheet 7 (Jurisdictional): DATA_START_ROW=6, cols A-T (20)
      Sheet 9 (Evidence): DATA_START_ROW=5, cols A-I (9)
    """
    logger.info("Populating S3 - Secure Configuration...")

    def _base_cols(r, ws, svc, config_item, status=STATUS_COMPLIANT, team='DevOpsSec'):
        """Write base columns A-Q for S3 assessment sheets."""
        ws.cell(row=r, column=2, value=svc['name'])           # B: Cloud_Service_Name
        ws.cell(row=r, column=3, value=svc['provider'])        # C: Vendor_Name
        ws.cell(row=r, column=4, value=svc['type'])            # D: Service_Type
        ws.cell(row=r, column=5, value='Production')           # E: Environment
        ws.cell(row=r, column=6, value=svc['criticality'])     # F: Service_Criticality
        ws.cell(row=r, column=7, value=svc['data_class'])      # G: Data_Classification
        ws.cell(row=r, column=8, value=config_item)            # H: Configuration_Item
        ws.cell(row=r, column=9, value=status)                 # I: Status
        ws.cell(row=r, column=10, value='/evidence/s3/configs/')  # J: Evidence_Location
        ws.cell(row=r, column=16, value=team)                  # P: Responsible_Team

    # --- Sheet 2: Configuration Baseline (row 5, A-Q + R-AB, 28 cols) ---
    ws = wb['2. Configuration Baseline']
    configs = [
        (SAMPLE_SERVICES[0], 'MFA Conditional Access Policy', STATUS_COMPLIANT, 'v2.3', 'CHG-2025-042', 'Yes', 'Level 1', 'Yes', 'Yes', 'Yes', 'Yes', 'Full Compliance', 'Compliant', 'No AI Systems'),
        (SAMPLE_SERVICES[1], 'Azure NSG Rules Baseline', STATUS_COMPLIANT, 'v1.8', 'CHG-2025-038', 'Yes', 'Level 2', 'Yes', 'Yes', 'Yes', 'Yes', 'Full Compliance', 'Compliant', 'No AI Systems'),
        (SAMPLE_SERVICES[2], 'S3 Bucket Encryption Policy', STATUS_COMPLIANT, 'v1.2', 'CHG-2025-015', 'Yes', 'Level 1', 'Yes', 'Yes', 'Yes', 'Yes', 'Full Compliance', 'Compliant', 'No AI Systems'),
        (SAMPLE_SERVICES[3], 'SSO/SAML Configuration', STATUS_PARTIAL, 'v1.0', 'CHG-2025-051', 'Partial', 'N/A', 'Yes', 'Yes', 'No', 'Yes', 'Partial Compliance', 'Partial (In Progress)', 'No AI Systems'),
        (SAMPLE_SERVICES[4], 'ITSM Access Policies', STATUS_COMPLIANT, 'v3.1', 'CHG-2025-029', 'Yes', 'Level 1', 'Yes', 'Yes', 'Yes', 'Yes', 'Full Compliance', 'Compliant', 'No AI Systems'),
    ]
    for i, (svc, item, status, ver, chg, drift, cis, validated, backup, priv, sec_default, dora, nis2, ai) in enumerate(configs):
        r = 5 + i
        _base_cols(r, ws, svc, item, status)
        ws.cell(row=r, column=18, value=ver)       # R: Baseline_Version
        ws.cell(row=r, column=19, value=chg)       # S: Change_Mgmt_ID
        ws.cell(row=r, column=20, value=drift)     # T: Drift_Monitoring
        ws.cell(row=r, column=21, value=cis)       # U: CIS_Benchmark
        ws.cell(row=r, column=22, value=TODAY_STR)  # V: Last_Validated
        ws.cell(row=r, column=23, value=backup)    # W: Config_Backup
        ws.cell(row=r, column=24, value=priv)      # X: Privileged_Access_Log
        ws.cell(row=r, column=25, value=sec_default)  # Y: Security_By_Default
        ws.cell(row=r, column=26, value=dora)      # Z: DORA_Config_Documentation
        ws.cell(row=r, column=27, value=nis2)      # AA: NIS2_Secure_Deployment
        ws.cell(row=r, column=28, value=ai)        # AB: AI_System_Deployment_Controls
    logger.info(f"   [OK] Configuration Baseline: {len(configs)} configs")

    # --- Sheet 3: Access Control Setup (row 5, A-Q + R-X, 24 cols) ---
    ws = wb['3. Access Control Setup']
    for i, svc in enumerate(SAMPLE_SERVICES[:5]):
        r = 5 + i
        _base_cols(r, ws, svc, 'Identity & Access Configuration', STATUS_COMPLIANT, 'Security')
        ws.cell(row=r, column=18, value='Yes')              # R: SSO_Integrated
        ws.cell(row=r, column=19, value='Yes (All Users)')  # S: MFA_Enforced
        ws.cell(row=r, column=20, value='Yes')              # T: RBAC_Implemented
        ws.cell(row=r, column=21, value='Yes')              # U: Privileged_Access_JIT
        ws.cell(row=r, column=22, value='Yes')              # V: Service_Accounts_Inventoried
        ws.cell(row=r, column=23, value=TODAY_STR)           # W: Last_Access_Review
        ws.cell(row=r, column=24, value=3)                   # X: Admin_Account_Count
    logger.info("   [OK] Access Control Setup: 5 services")

    # --- Sheet 4: Network Security (row 5, A-Q + R-X, 24 cols) ---
    ws = wb['4. Network Security']
    for i, svc in enumerate(SAMPLE_SERVICES[:5]):
        r = 5 + i
        _base_cols(r, ws, svc, 'Network Configuration', STATUS_COMPLIANT, 'Cloud Teams')
        ws.cell(row=r, column=18, value='Yes (Allowlist)')     # R: IP_Restrictions
        ws.cell(row=r, column=19, value='Yes (Private Link)')  # S: Private_Connectivity
        ws.cell(row=r, column=20, value='Yes')                 # T: Network_Segmentation
        ws.cell(row=r, column=21, value='Yes')                 # U: WAF_Enabled
        ws.cell(row=r, column=22, value='Yes (Advanced)')      # V: DDoS_Protection
        ws.cell(row=r, column=23, value='Yes')                 # W: Firewall_Rules_Documented
        ws.cell(row=r, column=24, value=2)                     # X: Public_Endpoints_Count
    logger.info("   [OK] Network Security: 5 services")

    # --- Sheet 5: Encryption Configuration (row 5, A-Q + R-X, 24 cols) ---
    ws = wb['5. Encryption Configuration']
    for i, svc in enumerate(SAMPLE_SERVICES[:5]):
        r = 5 + i
        cmk = svc['data_class'] == 'Restricted'
        _base_cols(r, ws, svc, 'Encryption Settings', STATUS_COMPLIANT, 'Security')
        ws.cell(row=r, column=18, value='Yes (CMK)' if cmk else 'Yes (Provider Key)')  # R
        ws.cell(row=r, column=19, value='Yes (TLS 1.3)')       # S: Encryption_In_Transit
        ws.cell(row=r, column=20, value='AES-256')              # T: Encryption_Algorithm
        ws.cell(row=r, column=21, value='Customer Managed (HSM)' if cmk else 'Provider Managed')  # U
        ws.cell(row=r, column=22, value='90 days')              # V: Key_Rotation_Period
        ws.cell(row=r, column=23, value='Yes' if cmk else 'No')  # W: HSM_Used
        ws.cell(row=r, column=24, value=TODAY_STR)               # X: Last_Key_Rotation
    logger.info("   [OK] Encryption Configuration: 5 services")

    # --- Sheet 6: Deployment Checklist (row 5, A-Q + R-X, 24 cols) ---
    ws = wb['6. Deployment Checklist']
    for i, svc in enumerate(SAMPLE_SERVICES[:5]):
        r = 5 + i
        _base_cols(r, ws, svc, 'Deployment Verification', STATUS_COMPLIANT, 'DevOps')
        ws.cell(row=r, column=18, value='Yes')      # R: Pre_Deploy_Scan
        ws.cell(row=r, column=19, value='Yes')      # S: Security_Approved
        ws.cell(row=r, column=20, value='Yes')      # T: Backup_Tested
        ws.cell(row=r, column=21, value='Yes')      # U: Monitoring_Enabled
        ws.cell(row=r, column=22, value='Yes')      # V: Runbook_Available
        ws.cell(row=r, column=23, value='Yes')      # W: Rollback_Plan
        ws.cell(row=r, column=24, value=TODAY_STR)   # X: Deployment_Date
    logger.info("   [OK] Deployment Checklist: 5 services")

    # --- Sheet 7: Jurisdictional Risk (row 6, cols A-T, 20 cols) ---
    ws = wb['7. Jurisdictional Risk']
    for i, svc in enumerate(SAMPLE_SERVICES):
        r = 6 + i
        us_hq = svc['hq_jurisdiction'] == 'United States'
        ws.cell(row=r, column=2, value=svc['name'])
        ws.cell(row=r, column=3, value=svc['provider'])
        ws.cell(row=r, column=4, value='United States' if us_hq else 'Germany')
        ws.cell(row=r, column=5, value=svc['hq_jurisdiction'])
        ws.cell(row=r, column=6, value='Yes' if us_hq else 'No')
        ws.cell(row=r, column=7, value=svc['cloud_act'])
        ws.cell(row=r, column=8, value=svc['region'])
        ws.cell(row=r, column=9, value=svc['eu_boundary'])
        ws.cell(row=r, column=10, value='Yes' if svc['data_class'] == 'Restricted' else 'No')
        ws.cell(row=r, column=14, value='Medium' if us_hq else 'Low')
        ws.cell(row=r, column=15, value='CISO')
        ws.cell(row=r, column=16, value=TODAY_STR)
    logger.info(f"   [OK] Jurisdictional Risk: {len(SAMPLE_SERVICES)} entries")

    # --- Sheet 9: Evidence Register (row 5, cols A-I, 9 cols) ---
    ws = wb['9. Evidence Register']
    evidence = [
        ('EV-SCD-001', 'Microsoft 365', 'Configuration Baseline', 'Screenshot', 'MFA policy screenshot from Azure AD', '/evidence/s3/azure_mfa.png', TODAY_STR, 'Security', 'Current'),
        ('EV-SCD-002', 'Azure Cloud Platform', 'Network Security', 'Config Export', 'NSG rules export from Azure Portal', '/evidence/s3/azure_nsg.json', TODAY_STR, 'Cloud Teams', 'Current'),
        ('EV-SCD-003', 'AWS S3 Storage', 'Encryption', 'Scan Report', 'S3 encryption scan results', '/evidence/s3/aws_encryption_scan.pdf', TODAY_STR, 'Security', 'Current'),
        ('EV-SCD-004', 'Salesforce CRM', 'Access Control', 'Screenshot', 'SSO configuration screenshot', '/evidence/s3/sf_sso.png', TODAY_STR, 'IT Operations', 'Current'),
        ('EV-SCD-005', 'ServiceNow ITSM', 'Deployment', 'Test Result', 'Deployment checklist verification', '/evidence/s3/snow_deploy_test.pdf', TODAY_STR, 'DevOps', 'Current'),
    ]
    for i, ev in enumerate(evidence):
        r = 5 + i
        for j, val in enumerate(ev):
            ws.cell(row=r, column=j + 1, value=val)
    logger.info(f"   [OK] Evidence Register: {len(evidence)} entries")

    return wb


# =============================================================================
# S4 - GOVERNANCE POPULATION
# =============================================================================

def populate_s4_governance(wb):
    """Populate S4 Governance workbook.

    Sheet structure (from generator):
      Sheets 2-6: DATA_START_ROW=6 (row 5 = example), base cols A-Q + extended R+
      Sheet 7 (Exit Review): DATA_START_ROW=4, cols A-N (14)
      Sheet 8 (Jurisdictional): DATA_START_ROW=6, cols A-T (20)
      Sheet 10 (Evidence): DATA_START_ROW=5, cols A-I (9)
    """
    logger.info("Populating S4 - Governance...")

    def _s4_base(r, ws, svc, col_g_val, status=STATUS_COMPLIANT, team='IT Ops'):
        """Write base columns A-Q for S4 assessment sheets."""
        ws.cell(row=r, column=1, value=svc['name'])            # A: Cloud Service Name
        ws.cell(row=r, column=2, value=svc['provider'])         # B: Vendor Name
        ws.cell(row=r, column=3, value=svc['type'])             # C: Service Type
        ws.cell(row=r, column=4, value='Q4')                    # D: Review Period
        ws.cell(row=r, column=5, value=svc['criticality'])      # E: Service Criticality
        ws.cell(row=r, column=6, value=svc['data_class'])       # F: Data Classification
        ws.cell(row=r, column=7, value=col_g_val)               # G: varies per sheet
        ws.cell(row=r, column=8, value=status)                  # H: Status
        ws.cell(row=r, column=9, value='/evidence/s4/')         # I: Evidence Location
        ws.cell(row=r, column=15, value=team)                   # O: Responsible Team
        ws.cell(row=r, column=17, value=TODAY_STR)              # Q: Last Review Date

    # --- Sheet 2: Access Review (row 6, A-Q + R-X, 24 cols) ---
    ws = wb['2. Access Review']
    for i, svc in enumerate(SAMPLE_SERVICES[:5]):
        r = 6 + i
        _s4_base(r, ws, svc, 'Full Recertification', STATUS_COMPLIANT, 'Security')
        ws.cell(row=r, column=18, value=(TODAY - timedelta(days=90)).strftime('%Y-%m-%d'))  # R: Last Review Date
        ws.cell(row=r, column=19, value='Passed')              # S: Review Outcome
        ws.cell(row=r, column=20, value=svc['users'])          # T: Total Accounts
        ws.cell(row=r, column=21, value=2)                     # U: Orphan Accounts
        ws.cell(row=r, column=22, value=3)                     # V: Excessive Privileges
        ws.cell(row=r, column=23, value=5)                     # W: Accounts Remediated
        ws.cell(row=r, column=24, value=(TODAY + timedelta(days=90)).strftime('%Y-%m-%d'))  # X: Next Review
    logger.info("   [OK] Access Review: 5 services")

    # --- Sheet 3: Change Management (row 6, A-Q + R-X, 24 cols) ---
    ws = wb['3. Change Management']
    for i, svc in enumerate(SAMPLE_SERVICES[:5]):
        r = 6 + i
        _s4_base(r, ws, svc, 'Provider Change', STATUS_COMPLIANT, 'IT Ops')
        ws.cell(row=r, column=18, value=3)                     # R: Change Count
        ws.cell(row=r, column=19, value='Medium')              # S: Impact Level
        ws.cell(row=r, column=20, value='Approved')            # T: Approval Status
        ws.cell(row=r, column=21, value='Yes')                 # U: Rollback Plan
        ws.cell(row=r, column=22, value='Yes')                 # V: Rollback Tested
        ws.cell(row=r, column=23, value='Yes')                 # W: Post-Change Review
        ws.cell(row=r, column=24, value='Yes')                 # X: Security Impact
    logger.info("   [OK] Change Management: 5 services")

    # --- Sheet 4: Incident Management (row 6, A-Q + R-AB, 28 cols) ---
    ws = wb['4. Incident Management']
    incidents = [
        (SAMPLE_SERVICES[0], 'P3-Medium', STATUS_COMPLIANT, 1, 2.5, 'Yes', 'Yes', 'No', 'No', 'Yes'),
        (SAMPLE_SERVICES[1], 'P2-High', STATUS_PARTIAL, 2, 4.0, 'Yes', 'Yes', 'Yes', 'Yes', 'Yes'),
        (SAMPLE_SERVICES[3], 'P4-Low', STATUS_COMPLIANT, 1, 1.0, 'Yes', 'No', 'No', 'No', 'Yes'),
    ]
    for i, (svc, severity, status, count, mttr, rca, playbook, notified, impact, lessons) in enumerate(incidents):
        r = 6 + i
        _s4_base(r, ws, svc, severity, status, 'Security')
        ws.cell(row=r, column=18, value=count)                  # R: Incident Count
        ws.cell(row=r, column=19, value=mttr)                   # S: MTTR (Hours)
        ws.cell(row=r, column=20, value=rca)                    # T: Root Cause Documented
        ws.cell(row=r, column=21, value=playbook)               # U: Playbook Updated
        ws.cell(row=r, column=22, value=notified)               # V: Vendor Notified
        ws.cell(row=r, column=23, value=impact)                 # W: Customer Impact
        ws.cell(row=r, column=24, value=lessons)                # X: Lessons Learned
        ws.cell(row=r, column=25, value='Continuous Monitoring')  # Y: DORA
        ws.cell(row=r, column=26, value='Minor (No notification required)')  # Z: NIS2
        ws.cell(row=r, column=27, value='No Monitoring Required')  # AA: AI
        ws.cell(row=r, column=28, value='CISO')                 # AB: Regulatory Owner
    logger.info(f"   [OK] Incident Management: {len(incidents)} incidents")

    # --- Sheet 5: Business Continuity (row 6, A-Q + R-X, 24 cols) ---
    ws = wb['5. Business Continuity']
    for i, svc in enumerate(SAMPLE_SERVICES[:5]):
        r = 6 + i
        tier = 'Tier 1 (<4hr)' if svc['criticality'] == 'Critical' else 'Tier 2 (<24hr)'
        rto_target = 4 if svc['criticality'] == 'Critical' else 24
        _s4_base(r, ws, svc, tier, STATUS_COMPLIANT, 'IT Ops')
        ws.cell(row=r, column=18, value=(TODAY - timedelta(days=90)).strftime('%Y-%m-%d'))  # R: Last Test
        ws.cell(row=r, column=19, value='Passed')               # S: Test Result
        ws.cell(row=r, column=20, value=rto_target)              # T: RTO Target
        ws.cell(row=r, column=21, value=rto_target - 1)          # U: RTO Achieved
        ws.cell(row=r, column=22, value=rto_target // 4 or 1)    # V: RPO Target
        ws.cell(row=r, column=23, value=max(1, rto_target // 4 - 1))  # W: RPO Achieved
        ws.cell(row=r, column=24, value=(TODAY + timedelta(days=90)).strftime('%Y-%m-%d'))  # X: Next Test
    logger.info("   [OK] Business Continuity: 5 services")

    # --- Sheet 6: Vendor Risk Monitoring (row 6, A-Q + R-X, 24 cols) ---
    ws = wb['6. Vendor Risk Monitoring']
    for i, svc in enumerate(SAMPLE_SERVICES[:5]):
        r = 6 + i
        risk = 'Medium' if svc['criticality'] in ['Critical', 'High'] else 'Low'
        _s4_base(r, ws, svc, risk, STATUS_COMPLIANT, 'Risk Management')
        ws.cell(row=r, column=18, value='Stable')               # R: Risk Score Trend
        ws.cell(row=r, column=19, value=1)                      # S: Security Incidents YTD
        ws.cell(row=r, column=20, value='Yes')                  # T: Cert Expiry Tracked
        ws.cell(row=r, column=21, value='2026-06-15')           # U: Next Cert Expiry
        ws.cell(row=r, column=22, value='Strong')               # V: Financial Health
        ws.cell(row=r, column=23, value=TODAY_STR)              # W: Last Assessment Date
        ws.cell(row=r, column=24, value='No')                   # X: Reassessment Trigger
    logger.info("   [OK] Vendor Risk Monitoring: 5 services")

    # --- Sheet 7: Exit Strategy Review (row 4, cols A-N, 14 cols) ---
    ws = wb['7. Exit Strategy Review']
    for i, svc in enumerate(SAMPLE_SERVICES):
        r = 4 + i
        last_review = TODAY - timedelta(days=180)
        next_review = TODAY + timedelta(days=185)
        poc_req = 'Yes (Critical)' if svc['criticality'] == 'Critical' else 'No (Not Critical)'
        ws.cell(row=r, column=1, value=svc['name'])              # A: Service Name
        ws.cell(row=r, column=2, value=svc['provider'])           # B: Provider
        ws.cell(row=r, column=3, value=svc['criticality'])        # C: Criticality
        ws.cell(row=r, column=4, value=svc['exit_strategy'])      # D: Exit Strategy Type
        ws.cell(row=r, column=5, value=last_review.strftime('%Y-%m-%d'))  # E: Last Review
        ws.cell(row=r, column=6, value=next_review.strftime('%Y-%m-%d'))  # F: Next Review
        ws.cell(row=r, column=7, value='Current')                 # G: Review Status
        ws.cell(row=r, column=8, value=poc_req)                   # H: PoC Testing Required
        if svc['criticality'] == 'Critical':
            ws.cell(row=r, column=9, value=(TODAY - timedelta(days=60)).strftime('%Y-%m-%d'))  # I
            ws.cell(row=r, column=10, value='Pass')               # J: PoC Test Result
            ws.cell(row=r, column=11, value=(TODAY + timedelta(days=305)).strftime('%Y-%m-%d'))  # K
        else:
            ws.cell(row=r, column=10, value='Not Applicable')
        ws.cell(row=r, column=12, value='No')                     # L: Exit Strategy Changed
        ws.cell(row=r, column=13, value='J. Smith')               # M: Reviewer Name
        ws.cell(row=r, column=14, value='Annual review completed')  # N: Notes
    logger.info(f"   [OK] Exit Strategy Review: {len(SAMPLE_SERVICES)} services")

    # --- Sheet 8: Jurisdictional Risk (row 6, cols A-T, 20 cols) ---
    ws = wb['8. Jurisdictional Risk']
    for i, svc in enumerate(SAMPLE_SERVICES):
        r = 6 + i
        us_hq = svc['hq_jurisdiction'] == 'United States'
        ws.cell(row=r, column=2, value=svc['name'])
        ws.cell(row=r, column=3, value=svc['provider'])
        ws.cell(row=r, column=4, value='United States' if us_hq else 'Germany')
        ws.cell(row=r, column=5, value=svc['hq_jurisdiction'])
        ws.cell(row=r, column=6, value='Yes' if us_hq else 'No')
        ws.cell(row=r, column=7, value=svc['cloud_act'])
        ws.cell(row=r, column=8, value=svc['region'])
        ws.cell(row=r, column=9, value=svc['eu_boundary'])
        ws.cell(row=r, column=10, value='Yes' if svc['data_class'] == 'Restricted' else 'No')
        ws.cell(row=r, column=14, value='Medium' if us_hq else 'Low')
        ws.cell(row=r, column=15, value='CISO')
        ws.cell(row=r, column=16, value=TODAY_STR)
    logger.info(f"   [OK] Jurisdictional Risk: {len(SAMPLE_SERVICES)} entries")

    # --- Sheet 10: Evidence Register (row 5, cols A-I, 9 cols) ---
    ws = wb['10. Evidence Register']
    evidence = [
        ('EV-OGR-001', 'Access Review', 'Microsoft 365', 'Access review report', 'Q4 access recertification for M365', '/evidence/s4/m365_access_review.pdf', TODAY_STR, 'Security', 'Verified'),
        ('EV-OGR-002', 'Change Management', 'Azure Cloud Platform', 'Change ticket', 'Azure NSG change ticket CHG-2025-038', '/evidence/s4/azure_chg_038.pdf', TODAY_STR, 'IT Operations', 'Verified'),
        ('EV-OGR-003', 'Incident Management', 'Azure Cloud Platform', 'Incident report', 'P2 incident — S3 misconfiguration', '/evidence/s4/inc_002_report.pdf', TODAY_STR, 'Security', 'Verified'),
        ('EV-OGR-004', 'Business Continuity', 'SAP S/4HANA Cloud', 'BC/DR test report', 'Annual BC/DR test results for SAP', '/evidence/s4/sap_bcdr_test.pdf', TODAY_STR, 'IT Operations', 'Verified'),
        ('EV-OGR-005', 'Vendor Risk Monitoring', 'Crowdstrike Falcon', 'Risk assessment', 'Annual vendor risk assessment', '/evidence/s4/cs_risk_assessment.pdf', TODAY_STR, 'Risk Management', 'Verified'),
    ]
    for i, ev in enumerate(evidence):
        r = 5 + i
        for j, val in enumerate(ev):
            ws.cell(row=r, column=j + 1, value=val)
    logger.info(f"   [OK] Evidence Register: {len(evidence)} entries")

    return wb


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def find_latest_workbook(base_dir, pattern):
    """Find the most recent workbook matching a glob pattern."""
    matches = sorted(glob.glob(os.path.join(base_dir, pattern)), key=os.path.getmtime, reverse=True)
    return matches[0] if matches else None


def main():
    """Main execution - populate all workbooks with sample data."""
    logger.info("=" * 70)
    logger.info("ISMS-IMP-A.5.23 - Sample Data Population Script")
    logger.info("=" * 70)

    # Resolve workbook directory relative to this script
    script_dir = Path(__file__).resolve().parent
    base_path = str(script_dir.parent.parent / "WKBK" / "90_workbooks")
    output_path = os.path.join(base_path, "populated")

    # Create output directory
    os.makedirs(output_path, exist_ok=True)

    # Workbook patterns (glob-based discovery, no hardcoded dates)
    workbook_patterns = {
        'S1': ('ISMS-IMP-A.5.23.S1_Inventory_*.xlsx', populate_s1_inventory),
        'S2': ('ISMS-IMP-A.5.23.S2_VendorDD_*.xlsx', populate_s2_vendor_dd),
        'S3': ('ISMS-IMP-A.5.23.S3_SecureConfig_*.xlsx', populate_s3_secure_config),
        'S4': ('ISMS-IMP-A.5.23.S4_Governance_*.xlsx', populate_s4_governance),
    }

    populated_count = 0
    for wb_id, (pattern, populate_func) in workbook_patterns.items():
        filepath = find_latest_workbook(base_path, pattern)

        if not filepath:
            logger.warning(f"[SKIP] No workbook found matching {pattern}")
            continue

        filename = os.path.basename(filepath)
        logger.info(f"\n[{wb_id}] Loading {filename}...")
        wb = load_workbook(filepath)

        # Populate with sample data
        wb = populate_func(wb)

        # Save to output folder
        output_file = os.path.join(output_path, filename.replace(".xlsx", "_SAMPLE.xlsx"))
        wb.save(output_file)
        logger.info(f"[{wb_id}] Saved: {os.path.basename(output_file)}")
        populated_count += 1

    # Summary
    logger.info("\n" + "=" * 70)
    logger.info("SAMPLE DATA POPULATION COMPLETE")
    logger.info("=" * 70)
    logger.info(f"  Workbooks populated: {populated_count}/4")
    logger.info(f"  Output folder: {output_path}")
    logger.info("\nNote: S5 Dashboard consolidates from S1-S4 data.")
    logger.info("  Next steps:")
    logger.info("  1. Run normalize_assessment_files_reg_a523.py")
    logger.info("  2. Run consolidate_reg_a523_dashboard.py")


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-02-10
# QA_STATUS: PASSED - UTILITY SCRIPT
# CHANGES: Complete rewrite — correct column mappings from generator specs,
#          correct start rows, emoji status values, evidence register population,
#          all 4 workbooks (S1-S4) with all assessment sheets populated
# =============================================================================
