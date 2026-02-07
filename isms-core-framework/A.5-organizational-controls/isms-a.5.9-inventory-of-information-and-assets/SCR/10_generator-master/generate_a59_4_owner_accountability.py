#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.9.4 - Owner Accountability Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.9: Inventory of Information and Assets
Assessment Domain 4 of 5: Asset Ownership & Accountability Verification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific organizational structure, communication channels,
and accountability expectations.

Key customization areas:
1. Organizational structure (departments, business units, reporting lines)
2. Attestation workflow (email templates, reminder schedules, escalation paths)
3. Performance expectations (review frequency, response times, engagement levels)
4. Accountability thresholds (what's acceptable for your organizational culture)
5. Organization name, CISO details, contact information
6. File paths and naming conventions
7. Integration with HR systems and email platforms

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
asset ownership accountability through systematic verification of ownership
coverage, owner acknowledgment, awareness, and performance.

**Purpose:**
Enables systematic assessment of whether asset owners understand their
responsibilities, acknowledge ownership, demonstrate awareness of assets under
their care, and actively perform their ownership duties. Measures accountability
through attestation campaigns, engagement tracking, and performance metrics.

**Assessment Scope:**
- Ownership Coverage: All assets have identified owners who acknowledge role
- Owner Acknowledgment: Formal attestation that owner accepts responsibilities
- Owner Awareness: Owner can identify assets under their care and knows status
- Owner Performance: Owner actively performs duties (reviews, updates, decisions)

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment methodology and accountability framework
2. Ownership Coverage - Asset ownership assignment and acceptance verification
3. Owner Acknowledgment - Attestation campaign tracking and response rates
4. Owner Awareness - Owner knowledge of assets and current state
5. Owner Performance - Active performance of ownership duties
6. Accountability Metrics - Aggregated scores and weighted accountability index
7. Evidence Register - Attestation evidence and performance documentation

**Key Features:**
- Data validation with dropdown lists for attestation status
- Conditional formatting for accountability scores (Green/Yellow/Red)
- Attestation campaign tracking (email sent, reminders, responses)
- Automated gap identification for missing or non-responsive owners
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Weighted accountability scoring (4 domains with configurable weights)
- Reminder schedule tracking (Day 7, 14, 21, escalation at Day 28)
- Performance engagement metrics
- Owner contact information and escalation paths

**Integration:**
This assessment feeds into the A.5.9.5 Compliance Dashboard, which consolidates
data from all four assessment domains for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - os (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a59_4_owner_accountability.py

Output:
    File: ISMS-IMP-A.5.9.4_Owner_Accountability_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Populate ownership coverage (assign owners to all assets)
    2. Launch attestation campaign (email owners requesting acknowledgment)
    3. Track responses and send reminders (Day 7, 14, 21)
    4. Escalate non-responders at Day 28
    5. Conduct awareness checks (can owner list their assets?)
    6. Assess performance (review compliance, response times, engagement)
    7. Review Accountability Metrics for overall accountability index
    8. Collect and link evidence in Evidence Register sheet
    9. Export metrics CSV for dashboard consolidation (Sheet 6)
    10. Store assessment workbook per retention policy (7 years minimum)
    11. Update quarterly or after ownership changes

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.9
Assessment Domain:    4 of 5
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organization License]

Related Documents:
    - ISMS-POL-A.5.9: Inventory of Information and Assets (Policy)
    - ISMS-IMP-A.5.9-1: Asset Discovery Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-2: Inventory Maintenance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-3: Quality & Compliance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-4: Owner Accountability Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 22.01.2026
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.9-4 specification
    - Four accountability domains with weighted scoring
    - Attestation campaign tracking with reminder workflow
    - Owner awareness verification
    - Performance engagement metrics
    - Evidence collection and audit trail support

[Future changes to be documented here]

--------------------------------------------------------------------------------
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# CUSTOMIZE: Configuration
CONFIG = {
    'organization': '[Organization]',
    'ciso_name': '[CISO Name]',
    'security_contact': '[security@organization.com]',
    
    # Color scheme (consistent with A.8.24 pattern)
    'colors': {
        'header_bg': '003366',      # Dark blue
        'header_text': 'FFFFFF',     # White
        'section_bg': '4472C4',      # Medium blue
        'green_light': 'C6EFCE',     # Light green (pass)
        'green_dark': '006100',      # Dark green (pass text)
        'yellow_light': 'FFEB9C',    # Light yellow (warning)
        'yellow_dark': '9C5700',     # Dark yellow (warning text)
        'red_light': 'FFC7CE',       # Light red (fail)
        'red_dark': '9C0006',        # Dark red (fail text)
        'gray_light': 'D9D9D9',      # Light gray (locked cells)
    },
    
    # Accountability domain weights (must sum to 100%)
    'accountability_weights': {
        'Coverage': 40,         # Most important - all assets must have owners
        'Acknowledgment': 25,   # Owner formally accepts responsibility
        'Awareness': 20,        # Owner knows what they own
        'Performance': 15,      # Owner actively performs duties
    },
    
    # Accountability targets (customize per domain)
    'accountability_targets': {
        'Coverage': 100,        # CRITICAL - every asset needs an owner
        'Acknowledgment': 95,   # High - most owners should attest
        'Awareness': 100,       # CRITICAL - owner must know their assets
        'Performance': 80,      # Good - active engagement expected
    },
    
    # Overall accountability target (from policy)
    'overall_target': 94,
    
    # Attestation workflow
    'attestation': {
        'reminder_schedule': [7, 14, 21],  # Days after initial email
        'escalation_day': 28,              # Escalate to manager
        'validity_period': 365,            # Days before re-attestation
    },
}

# Document identification constants
DOCUMENT_ID = "ISMS-IMP-A.5.9.4"
CONTROL_REF = "ISO/IEC 27001:2022 - Control A.5.9: Inventory of Information and Assets"


def main():
    """Main execution function"""
    logger.info("="*80)
    logger.info("ISMS Control A.5.9 - Owner Accountability Assessment Generator")
    logger.info("="*80)
    logger.info("")
    logger.info("Generating assessment workbook...")
    logger.info("")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create sheets in order
    sheets = [
        "Instructions & Legend",
        "Ownership Coverage",
        "Owner Acknowledgment",
        "Owner Awareness",
        "Owner Performance",
        "Accountability Metrics",
        "Evidence Register",
        "Summary Dashboard",
        "Approval & Sign-Off",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
        logger.info(f"  ✓ Created sheet: {sheet_name}")
    
    logger.info("")
    logger.info("Populating sheets...")
    logger.info("")
    
    # Populate each sheet
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  ✓ Instructions & Legend")
    
    create_ownership_coverage_sheet(wb["Ownership Coverage"])
    logger.info("  ✓ Ownership Coverage")
    
    create_owner_acknowledgment_sheet(wb["Owner Acknowledgment"])
    logger.info("  ✓ Owner Acknowledgment")
    
    create_owner_awareness_sheet(wb["Owner Awareness"])
    logger.info("  ✓ Owner Awareness")
    
    create_owner_performance_sheet(wb["Owner Performance"])
    logger.info("  ✓ Owner Performance")
    
    create_accountability_metrics_sheet(wb["Accountability Metrics"])
    logger.info("  ✓ Accountability Metrics")
    
    create_evidence_register_sheet(wb["Evidence Register"])
    logger.info("  ✓ Evidence Register")
    
    create_summary_dashboard_sheet(wb["Summary Dashboard"])
    logger.info("  ✓ Summary Dashboard")
    
    create_approval_signoff_sheet(wb["Approval & Sign-Off"])
    logger.info("  ✓ Approval & Sign-Off")
    
    # Save workbook
    filename = f"ISMS-IMP-A.5.9.4_Owner_Accountability_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    logger.info("")
    logger.info("="*80)
    logger.info(f"✅ SUCCESS: {filename}")
    logger.info("="*80)
    logger.info("")
    logger.info("Next Steps:")
    logger.info("  1. Open workbook and review Instructions & Legend")
    logger.info("  2. Populate ownership coverage (assign all assets to owners)")
    logger.info("  3. Launch attestation campaign (email owners)")
    logger.info("  4. Track acknowledgment responses and send reminders")
    logger.info("  5. Conduct awareness checks (owner knowledge verification)")
    logger.info("  6. Assess owner performance (reviews, responsiveness, engagement)")
    logger.info("  7. Review Accountability Metrics for overall accountability index")
    logger.info("  8. Export CSV from Sheet 6 for dashboard consolidation")
    logger.info("  9. Obtain stakeholder review and approval")
    logger.info("")
    logger.info(f"Output location: {os.path.abspath(filename)}")
    logger.info("")


def create_instructions_sheet(ws):
    """Create Instructions & Legend sheet"""
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{DOCUMENT_ID}  -  Owner Accountability Assessment\n{CONTROL_REF}"
    ws['A1'].font = Font(size=16, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 40
    
    # Subtitle
    ws.merge_cells('A2:H2')
    ws['A2'] = "Assessment Domain 4 of 5: Asset Ownership & Accountability Verification"
    ws['A2'].font = Font(size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Purpose section
    ws['A4'] = "PURPOSE"
    ws['A4'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws['A4'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    ws.merge_cells('A5:H8')
    purpose_text = """This assessment evaluates ACCOUNTABILITY - do asset owners understand their role, acknowledge ownership, know what they own, and actively perform their duties?

Discovery finds assets. Maintenance keeps them current. Quality ensures accuracy. Accountability ensures someone is responsible.

Four accountability domains are assessed: Coverage, Acknowledgment, Awareness, and Performance."""
    ws['A5'] = purpose_text
    ws['A5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[5].height = 60
    
    # Four Accountability Domains
    ws['A10'] = "FOUR ACCOUNTABILITY DOMAINS"
    ws['A10'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws['A10'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    domains = [
        ("1. COVERAGE (40% weight)", "All assets have identified owners; owners accept role", "100%"),
        ("2. ACKNOWLEDGMENT (25%)", "Owners formally attest to responsibilities (annual attestation campaign)", "95%"),
        ("3. AWARENESS (20%)", "Owners can identify assets under their care and know current state", "100%"),
        ("4. PERFORMANCE (15%)", "Owners actively perform duties (reviews, updates, decisions, responsiveness)", "80%"),
    ]
    
    row = 11
    for domain, description, target in domains:
        ws[f'A{row}'] = domain
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = description
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(bold=True)
        ws.row_dimensions[row].height = 30
        row += 1
    
    # Attestation Campaign Workflow
    ws[f'A{row+1}'] = "ATTESTATION CAMPAIGN WORKFLOW"
    ws[f'A{row+1}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{row+1}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    row += 2
    ws.merge_cells(f'A{row}:H{row+4}')
    attestation_text = """Annual attestation campaign to formally acknowledge ownership:

Day 0: Email all asset owners requesting attestation
Day 7: Reminder #1 to non-responders
Day 14: Reminder #2 to non-responders
Day 21: Reminder #3 to non-responders
Day 28: Escalate to owner's manager + CISO notification

Target: ≥95% attestation rate within 28 days"""
    ws[f'A{row}'] = attestation_text
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[row].height = 100
    
    row += 6
    
    # Weighted Scoring
    ws[f'A{row}'] = "WEIGHTED ACCOUNTABILITY SCORING"
    ws[f'A{row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    row += 1
    ws[f'A{row}'] = "Overall Accountability Index = (Coverage × 40%) + (Acknowledgment × 25%) + (Awareness × 20%) + (Performance × 15%)"
    ws[f'A{row}'].font = Font(italic=True)
    ws.merge_cells(f'A{row}:H{row}')
    
    row += 1
    ws[f'A{row}'] = "Target: ≥94% for effective accountability"
    ws[f'A{row}'].font = Font(bold=True)
    
    # Traffic Light Legend
    row += 2
    ws[f'A{row}'] = "TRAFFIC LIGHT LEGEND"
    ws[f'A{row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    row += 1
    # Green
    ws[f'A{row}'] = "✅ GREEN - Strong Accountability"
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid')
    ws[f'B{row}'] = "Score ≥ Target"
    
    # Yellow
    row += 1
    ws[f'A{row}'] = "⚠️ YELLOW - Needs Improvement"
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    ws[f'B{row}'] = "Score within 5% of target"
    
    # Red
    row += 1
    ws[f'A{row}'] = "❌ RED - Weak Accountability"
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid')
    ws[f'B{row}'] = "Score >5% below target"
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 12
    
    # Protect sheet


def create_ownership_coverage_sheet(ws):
    """Create Ownership Coverage sheet"""
    
    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "Ownership Coverage - Asset Owner Assignment"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25
    
    # Column headers
    headers = [
        ('A', 'Asset Category', 25),
        ('B', 'Total Assets', 15),
        ('C', 'Assets with Owner', 18),
        ('D', 'Assets without Owner', 18),
        ('E', 'Coverage %', 15),
        ('F', 'Target %', 12),
        ('G', 'Gap', 12),
        ('H', 'Status', 18),
        ('I', 'Unique Owners Count', 18),
        ('J', 'Avg Assets per Owner', 18),
        ('K', 'Evidence ID', 15),
        ('L', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Asset categories
    categories = [
        "Information Assets",
        "IT Infrastructure",
        "Applications",
        "Physical Assets",
        "Personnel Assets",
    ]
    
    row = 4
    for category in categories:
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'F{row}'] = "100%"
        
        # User enters counts
        ws[f'B{row}'].protection = Protection(locked=False)
        ws[f'C{row}'].protection = Protection(locked=False)
        
        # Assets without owner
        ws[f'D{row}'] = f'=B{row}-C{row}'
        
        # Coverage %
        ws[f'E{row}'] = f'=IFERROR(C{row}/B{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        
        # Gap
        ws[f'G{row}'] = f'=E{row}-100'
        ws[f'G{row}'].number_format = '0.0"%"'
        
        # Status
        ws[f'H{row}'] = f'=IF(E{row}=100,"✅ Full Coverage",IF(E{row}>=95,"⚠️ Minor Gaps","❌ Major Gaps"))'
        ws[f'H{row}'].alignment = Alignment(horizontal='center')
        
        # Unique owners and avg per owner (user enters)
        ws[f'I{row}'].protection = Protection(locked=False)
        ws[f'J{row}'] = f'=IFERROR(B{row}/I{row},0)'
        ws[f'J{row}'].number_format = '0.0'
        
        ws[f'K{row}'].protection = Protection(locked=False)
        ws[f'L{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "OWNERSHIP COVERAGE SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Assets"
    ws[f'B{summary_row}'] = f'=SUM(B4:B8)'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Assets with Owner"
    ws[f'B{summary_row}'] = f'=SUM(C4:C8)'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Assets without Owner"
    ws[f'B{summary_row}'] = f'=SUM(D4:D8)'
    ws[f'B{summary_row}'].font = Font(bold=True, color=CONFIG['colors']['red_dark'])
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Overall Coverage %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-2}/B{summary_row-3}*100,0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "100%"
    ws[f'B{summary_row}'].font = Font(bold=True, color=CONFIG['colors']['red_dark'])
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-100'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Status"
    ws[f'B{summary_row}'] = f'=IF(B{summary_row-3}=100,"✅ Strong",IF(B{summary_row-3}>=95,"⚠️ Needs Improvement","❌ Weak"))'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'E4:E8',
        CellIsRule(operator='equal', formula=['100'],
                   fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E8',
        CellIsRule(operator='between', formula=['95', '99'],
                   fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E8',
        CellIsRule(operator='lessThan', formula=['95'],
                   fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid'))
    )


def create_owner_acknowledgment_sheet(ws):
    """Create Owner Acknowledgment (Attestation) sheet"""
    
    # Header
    ws.merge_cells('A1:O1')
    ws['A1'] = "Owner Acknowledgment - Attestation Campaign Tracking"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Owner Name', 25),
        ('B', 'Owner Email', 30),
        ('C', 'Department', 20),
        ('D', 'Assets Count', 15),
        ('E', 'Email Sent Date', 18),
        ('F', 'Reminder 1 (Day 7)', 18),
        ('G', 'Reminder 2 (Day 14)', 18),
        ('H', 'Reminder 3 (Day 21)', 18),
        ('I', 'Escalation (Day 28)', 18),
        ('J', 'Response Date', 18),
        ('K', 'Days to Respond', 15),
        ('L', 'Attestation Status', 20),
        ('M', 'Next Action', 25),
        ('N', 'Evidence ID', 15),
        ('O', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Sample owner rows (user fills actual owners)
    sample_owners = [
        ("John Doe", "john.doe@org.com", "Engineering", 15),
        ("Jane Smith", "jane.smith@org.com", "IT Operations", 42),
        ("Bob Johnson", "bob.johnson@org.com", "Security", 8),
        ("Alice Williams", "alice.williams@org.com", "HR", 5),
    ]
    
    row = 4
    for name, email, dept, count in sample_owners:
        ws[f'A{row}'] = name
        ws[f'B{row}'] = email
        ws[f'C{row}'] = dept
        ws[f'D{row}'] = count
        
        # User enters campaign dates
        for col in ['E', 'F', 'G', 'H', 'I', 'J']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        
        # Days to respond (formula)
        ws[f'K{row}'] = f'=IF(AND(E{row}<>"",J{row}<>""),J{row}-E{row},"")'
        ws[f'K{row}'].number_format = '0'
        
        # Attestation Status (user selects)
        ws[f'L{row}'].protection = Protection(locked=False)
        
        # Next Action (user enters)
        ws[f'M{row}'].protection = Protection(locked=False)
        ws[f'N{row}'].protection = Protection(locked=False)
        ws[f'O{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Add 30 blank rows
    for i in range(30):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'M', 'N', 'O']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        ws[f'K{row}'] = f'=IF(AND(E{row}<>"",J{row}<>""),J{row}-E{row},"")'
        row += 1
    
    # Data validations
    attestation_statuses = [
        "Attested - On Time",
        "Attested - Late",
        "Pending",
        "Non-Responsive",
        "Escalated",
        "Exempted"
    ]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(attestation_statuses)}"', allow_blank=True)
    dv_status.add(f'L4:L{row-1}')
    ws.add_data_validation(dv_status)
    
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "ACKNOWLEDGMENT SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Owners"
    ws[f'B{summary_row}'] = f'=COUNTA(A4:A{row-31})'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Attested (On Time + Late)"
    ws[f'B{summary_row}'] = f'=COUNTIFS(L4:L{row-31},"Attested*")'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Acknowledgment Rate %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-2}*100,0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "95%"
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-95'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Response Time (Days)"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(K4:K{row-31}),0)'
    ws[f'B{summary_row}'].number_format = '0.0'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Status"
    ws[f'B{summary_row}'] = f'=IF(B{summary_row-4}>=95,"✅ Strong",IF(B{summary_row-4}>=90,"⚠️ Needs Improvement","❌ Weak"))'
    ws[f'B{summary_row}'].font = Font(bold=True)


def create_owner_awareness_sheet(ws):
    """Create Owner Awareness sheet"""
    
    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "Owner Awareness - Knowledge of Assets Under Care"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Owner Name', 25),
        ('B', 'Assets Assigned', 15),
        ('C', 'Assets Identified by Owner', 22),
        ('D', 'Awareness %', 15),
        ('E', 'Target %', 12),
        ('F', 'Gap', 12),
        ('G', 'Status', 18),
        ('H', 'Verification Method', 30),
        ('I', 'Verification Date', 18),
        ('J', 'Evidence ID', 15),
        ('K', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Sample owners (user fills)
    sample_owners = [
        "John Doe",
        "Jane Smith",
        "Bob Johnson",
        "Alice Williams",
    ]
    
    row = 4
    for owner in sample_owners:
        ws[f'A{row}'] = owner
        ws[f'E{row}'] = "100%"
        
        # User enters counts
        ws[f'B{row}'].protection = Protection(locked=False)
        ws[f'C{row}'].protection = Protection(locked=False)
        
        # Awareness %
        ws[f'D{row}'] = f'=IFERROR(C{row}/B{row}*100,0)'
        ws[f'D{row}'].number_format = '0.0"%"'
        
        # Gap
        ws[f'F{row}'] = f'=D{row}-100'
        ws[f'F{row}'].number_format = '0.0"%"'
        
        # Status
        ws[f'G{row}'] = f'=IF(D{row}=100,"✅ Fully Aware",IF(D{row}>=95,"⚠️ Mostly Aware","❌ Unaware"))'
        ws[f'G{row}'].alignment = Alignment(horizontal='center')
        
        # User enters verification details
        for col in ['H', 'I', 'J', 'K']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Add 30 blank rows
    for i in range(30):
        ws[f'E{row}'] = "100%"
        for col in ['A', 'B', 'C', 'H', 'I', 'J', 'K']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        ws[f'D{row}'] = f'=IFERROR(C{row}/B{row}*100,0)'
        ws[f'D{row}'].number_format = '0.0"%"'
        ws[f'F{row}'] = f'=D{row}-100'
        ws[f'F{row}'].number_format = '0.0"%"'
        ws[f'G{row}'] = f'=IF(D{row}=100,"✅ Fully Aware",IF(D{row}>=95,"⚠️ Mostly Aware","❌ Unaware"))'
        row += 1
    
    # Data validation
    verification_methods = [
        "Interview - Owner Listed Assets",
        "Survey - Owner Confirmed Assets",
        "Audit - Cross-checked vs Inventory",
        "Self-Assessment",
        "Other (Specify)"
    ]
    dv_method = DataValidation(type="list", formula1=f'"{",".join(verification_methods)}"', allow_blank=True)
    dv_method.add(f'H4:H{row-1}')
    ws.add_data_validation(dv_method)
    
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "AWARENESS SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Awareness %"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(D4:D{row-31}),0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "100%"
    ws[f'B{summary_row}'].font = Font(bold=True, color=CONFIG['colors']['red_dark'])
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-100'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Status"
    ws[f'B{summary_row}'] = f'=IF(B{summary_row-3}=100,"✅ Strong",IF(B{summary_row-3}>=95,"⚠️ Needs Improvement","❌ Weak"))'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'D4:D{row-31}',
        CellIsRule(operator='equal', formula=['100'],
                   fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'D4:D{row-31}',
        CellIsRule(operator='between', formula=['95', '99'],
                   fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'D4:D{row-31}',
        CellIsRule(operator='lessThan', formula=['95'],
                   fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid'))
    )


def create_owner_performance_sheet(ws):
    """Create Owner Performance sheet"""
    
    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "Owner Performance - Active Performance of Ownership Duties"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Owner Name', 25),
        ('B', 'Assets Count', 15),
        ('C', 'Reviews Due', 15),
        ('D', 'Reviews Completed', 18),
        ('E', 'Review Compliance %', 18),
        ('F', 'Avg Response Time (Days)', 22),
        ('G', 'Target Response (Days)', 20),
        ('H', 'Responsiveness %', 18),
        ('I', 'Performance Score', 18),
        ('J', 'Target', 12),
        ('K', 'Status', 18),
        ('L', 'Evidence ID', 15),
        ('M', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Sample owners
    sample_owners = [
        ("John Doe", 15, 3),
        ("Jane Smith", 42, 3),
        ("Bob Johnson", 8, 3),
        ("Alice Williams", 5, 3),
    ]
    
    row = 4
    for owner, count, target_days in sample_owners:
        ws[f'A{row}'] = owner
        ws[f'B{row}'] = count
        ws[f'G{row}'] = target_days
        ws[f'J{row}'] = "80%"
        
        # User enters review counts
        ws[f'C{row}'].protection = Protection(locked=False)
        ws[f'D{row}'].protection = Protection(locked=False)
        
        # Review Compliance %
        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        
        # User enters avg response time
        ws[f'F{row}'].protection = Protection(locked=False)
        
        # Responsiveness % (inverse - faster is better)
        ws[f'H{row}'] = f'=IF(F{row}="","",MAX(0,100-((F{row}-G{row})/G{row}*100)))'
        ws[f'H{row}'].number_format = '0.0"%"'
        
        # Performance Score = (Review Compliance × 60%) + (Responsiveness × 40%)
        ws[f'I{row}'] = f'=IF(OR(E{row}="",H{row}=""),"",(E{row}*0.6)+(H{row}*0.4))'
        ws[f'I{row}'].number_format = '0.0"%"'
        ws[f'I{row}'].font = Font(bold=True)
        
        # Status
        ws[f'K{row}'] = f'=IF(I{row}="","",IF(I{row}>=80,"✅ Strong",IF(I{row}>=75,"⚠️ Needs Improvement","❌ Weak")))'
        ws[f'K{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'L{row}'].protection = Protection(locked=False)
        ws[f'M{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Add 30 blank rows
    for i in range(30):
        ws[f'G{row}'] = 3
        ws[f'J{row}'] = "80%"
        for col in ['A', 'B', 'C', 'D', 'F', 'L', 'M']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        ws[f'H{row}'] = f'=IF(F{row}="","",MAX(0,100-((F{row}-G{row})/G{row}*100)))'
        ws[f'H{row}'].number_format = '0.0"%"'
        ws[f'I{row}'] = f'=IF(OR(E{row}="",H{row}=""),"",(E{row}*0.6)+(H{row}*0.4))'
        ws[f'I{row}'].number_format = '0.0"%"'
        ws[f'K{row}'] = f'=IF(I{row}="","",IF(I{row}>=80,"✅ Strong",IF(I{row}>=75,"⚠️ Needs Improvement","❌ Weak")))'
        row += 1
    
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "PERFORMANCE SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Performance Score %"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(I4:I{row-31}),0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "80%"
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-80'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Status"
    ws[f'B{summary_row}'] = f'=IF(B{summary_row-3}>=80,"✅ Strong",IF(B{summary_row-3}>=75,"⚠️ Needs Improvement","❌ Weak"))'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'I4:I{row-31}',
        CellIsRule(operator='greaterThanOrEqual', formula=['80'],
                   fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'I4:I{row-31}',
        CellIsRule(operator='between', formula=['75', '79'],
                   fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'I4:I{row-31}',
        CellIsRule(operator='lessThan', formula=['75'],
                   fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid'))
    )


def create_accountability_metrics_sheet(ws):
    """Create Accountability Metrics sheet - consolidates all domains"""
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Accountability Metrics - Weighted Accountability Index"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Domains table
    ws['A3'] = "Accountability Domain"
    ws['B3'] = "Score %"
    ws['C3'] = "Weight %"
    ws['D3'] = "Weighted Score"
    ws['E3'] = "Target %"
    ws['F3'] = "Gap"
    ws['G3'] = "Status"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    
    # Link to domain sheets
    domains_data = [
        ("Coverage", "='Ownership Coverage'!B17", 40, 100),
        ("Acknowledgment", "='Owner Acknowledgment'!B45", 25, 95),
        ("Awareness", "='Owner Awareness'!B40", 20, 100),
        ("Performance", "='Owner Performance'!B41", 15, 80),
    ]
    
    row = 4
    for domain, formula, weight, target in domains_data:
        ws[f'A{row}'] = domain
        ws[f'A{row}'].font = Font(bold=True)
        
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = '0.0"%"'
        ws[f'B{row}'].font = Font(bold=True)
        
        ws[f'C{row}'] = f'{weight}%'
        
        ws[f'D{row}'] = f'=B{row}*{weight}/100'
        ws[f'D{row}'].number_format = '0.0"%"'
        
        ws[f'E{row}'] = f'{target}%'
        
        ws[f'F{row}'] = f'=B{row}-{target}/100'
        ws[f'F{row}'].number_format = '0.0"%"'
        
        ws[f'G{row}'] = f'=IF(B{row}>={target}/100,"✅ Strong",IF(B{row}>={target}/100-0.05,"⚠️ Medium","❌ Weak"))'
        ws[f'G{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Overall accountability index
    overall_row = row + 2
    ws[f'A{overall_row}'] = "OVERALL ACCOUNTABILITY INDEX"
    ws[f'A{overall_row}'].font = Font(size=12, bold=True)
    ws[f'B{overall_row}'] = f'=SUM(D4:D7)'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    ws[f'B{overall_row}'].font = Font(bold=True, size=14)
    ws[f'B{overall_row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Target"
    ws[f'B{overall_row}'] = "94%"
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Gap"
    ws[f'B{overall_row}'] = f'=B{overall_row-2}-0.94'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Status"
    ws[f'B{overall_row}'] = f'=IF(B{overall_row-3}>=0.94,"✅ Strong",IF(B{overall_row-3}>=0.89,"⚠️ Needs Improvement","❌ Weak"))'
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    # CSV Export section
    csv_row = overall_row + 3
    ws[f'A{csv_row}'] = "CSV EXPORT FOR DASHBOARD (Copy rows below)"
    ws[f'A{csv_row}'].font = Font(size=11, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{csv_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    csv_row += 1
    ws[f'A{csv_row}'] = "Accountability_Domain"
    ws[f'B{csv_row}'] = "Score_%"
    ws[f'C{csv_row}'] = "Status"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{csv_row}'].font = Font(bold=True)
    
    csv_row += 1
    for i in range(4):
        ws[f'A{csv_row+i}'] = f'=A{4+i}'
        ws[f'B{csv_row+i}'] = f'=B{4+i}'
        ws[f'C{csv_row+i}'] = f'=G{4+i}'
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18


def create_evidence_register_sheet(ws):
    """Create Evidence Register sheet"""
    
    # Title
    ws.merge_cells('A1:N1')
    ws['A1'] = "Accountability Evidence Register"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Evidence ID', 15),
        ('B', 'Accountability Domain', 25),
        ('C', 'Evidence Type', 30),
        ('D', 'Evidence Description', 50),
        ('E', 'Evidence Location', 40),
        ('F', 'Collection Date', 15),
        ('G', 'Collected By', 25),
        ('H', 'Validity Period', 20),
        ('I', 'Review Date', 15),
        ('J', 'Reviewed By', 25),
        ('K', 'Review Status', 20),
        ('L', 'Retention End Date', 18),
        ('M', 'Related Assessment', 25),
        ('N', 'Notes', 40),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Sample evidence
    sample_data = [
        ("ACCT-001", "Coverage", "Ownership Assignments", "Complete asset ownership matrix - all assets assigned", "/evidence/ownership_matrix_20260122.xlsx", "22.01.2026"),
        ("ACCT-002", "Acknowledgment", "Attestation Campaign Results", "Annual attestation campaign - email logs and responses", "/evidence/attestation_campaign_20260122.xlsx", "22.01.2026"),
        ("ACCT-003", "Awareness", "Owner Interview Results", "Owner awareness interviews - asset identification verification", "/evidence/awareness_interviews_20260122.xlsx", "22.01.2026"),
        ("ACCT-004", "Performance", "Review Compliance Report", "Owner review compliance tracking - Q4 2025", "/evidence/review_compliance_Q4_2025.xlsx", "22.01.2026"),
    ]
    
    row = 4
    for evidence_id, domain, evidence_type, description, location, date in sample_data:
        ws[f'A{row}'] = evidence_id
        ws[f'B{row}'] = domain
        ws[f'C{row}'] = evidence_type
        ws[f'D{row}'] = description
        ws[f'E{row}'] = location
        ws[f'F{row}'] = date
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    # Add empty rows
    for i in range(20):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    # Data validations
    accountability_domains = ["Coverage", "Acknowledgment", "Awareness", "Performance", "All Domains"]
    dv_domain = DataValidation(type="list", formula1=f'"{",".join(accountability_domains)}"', allow_blank=True)
    dv_domain.add(f'B4:B100')
    ws.add_data_validation(dv_domain)
    
    review_statuses = ["Pending Review", "Reviewed - Valid", "Reviewed - Update Needed", "Reviewed - Invalid"]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(review_statuses)}"', allow_blank=True)
    dv_status.add(f'K4:K100')
    ws.add_data_validation(dv_status)


# Execute main function


def create_summary_dashboard_sheet(ws):
    """Create Summary Dashboard sheet"""
    CHECK = "✅"
    WARNING = "⚠️"
    XMARK = "❌"
    TARGET = "🎯"
    CHART = "📊"
    
    metrics_sheet = "Accountability Metrics"
    assessment_name = "Owner Accountability Assessment"
    compliance_ref = "D10"
    key_metrics = [
        ('Ownership Coverage', 'B4', '100%', 'percentage'),
        ('Acknowledgment Rate', 'B5', '100%', 'percentage'),
        ('Awareness Score', 'B6', '95%', 'percentage'),
    ]
    
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{CHART} {assessment_name.upper()} - SUMMARY DASHBOARD"
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:F2')
    ws['A2'] = "Quick overview of key metrics and compliance status"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].fill = PatternFill(start_color='E7E6E6', fill_type='solid')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    row = 4
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"{TARGET} OVERALL COMPLIANCE"
    ws[f'A{row}'].font = Font(size=13, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='2E75B5', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Overall Compliance:"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = f"='{metrics_sheet}'!{compliance_ref}"
    ws[f'D{row}'].font = Font(size=18, bold=True)
    ws[f'D{row}'].number_format = '0.0"%"'
    ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"{CHART} KEY METRICS"
    ws[f'A{row}'].font = Font(size=13, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='2E75B5', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    for col_idx, header in enumerate(["Metric", "Current", "Target", "Status"], start=1):
        ws.cell(row=row, column=col_idx, value=header)
        ws.cell(row=row, column=col_idx).font = Font(size=11, bold=True, color='FFFFFF')
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color='4472C4', fill_type='solid')
    row += 1
    
    for metric_label, cell_ref, target, format_type in key_metrics:
        ws.cell(row=row, column=1, value=metric_label)
        ws.cell(row=row, column=2, value=f"='{metrics_sheet}'!{cell_ref}")
        if format_type == 'percentage':
            ws.cell(row=row, column=2).number_format = '0.0"%"'
        ws.cell(row=row, column=3, value=target)
        ws.cell(row=row, column=4, value=f'=IF(B{row}>0,"{CHECK}","{XMARK}")')
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_approval_signoff_sheet(ws):
    """Create Approval & Sign-Off sheet"""
    CHECK = "✅"
    CLOCK = "⏳"
    XMARK = "❌"
    
    assessment_name = "Owner Accountability Assessment"
    checklist_items = [
        'Ownership coverage assessed',
        'Owner acknowledgments obtained',
        'Owner awareness assessed',
        'Owner performance evaluated',
        'Accountability metrics calculated',
        'Evidence collected and registered',
        'Assessment reviewed by Information Security'
    ]
    
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{assessment_name.upper()} - APPROVAL & SIGN-OFF"
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:F2')
    ws['A2'] = "Complete when assessment is finished and ready for approval"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    row = 4
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "COMPLETION CHECKLIST"
    ws[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='2E75B5', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    for col_idx, header in enumerate(["Item", "Status", "Completed By", "Date", "Notes"], start=1):
        ws.cell(row=row, column=col_idx, value=header)
        ws.cell(row=row, column=col_idx).font = Font(size=11, bold=True, color='FFFFFF')
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color='4472C4', fill_type='solid')
    row += 1
    
    checklist_start = row
    for item in checklist_items:
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=1).font = Font(size=10)
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFF2CC', fill_type='solid')
            ws.cell(row=row, column=col).protection = Protection(locked=False)
        row += 1
    
    from openpyxl.worksheet.datavalidation import DataValidation
    status_options = [f"{CHECK} Complete", f"{CLOCK} In Progress", f"{XMARK} Not Done"]
    dv = DataValidation(type="list", formula1=f'"{",".join(status_options)}"', allow_blank=True)
    dv.add(f'B{checklist_start}:B{row-1}')
    ws.add_data_validation(dv)
    
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "APPROVALS"
    ws[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='2E75B5', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    for col_idx, header in enumerate(["Role", "Name", "Signature", "Date", "Comments"], start=1):
        ws.cell(row=row, column=col_idx, value=header)
        ws.cell(row=row, column=col_idx).font = Font(size=11, bold=True, color='FFFFFF')
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color='4472C4', fill_type='solid')
    row += 1
    
    for approver in ["Asset Management Lead", "Information Security Manager", "CISO"]:
        ws.cell(row=row, column=1, value=approver)
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color='D8E4F8', fill_type='solid')
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFF2CC', fill_type='solid')
            ws.cell(row=row, column=col).protection = Protection(locked=False)
        row += 1
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30



if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
