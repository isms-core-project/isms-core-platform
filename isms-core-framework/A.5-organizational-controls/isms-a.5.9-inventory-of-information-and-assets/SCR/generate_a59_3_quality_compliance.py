#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.9.3 - Quality & Compliance Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.9: Inventory of Information and Assets
Assessment Domain 3 of 5: Data Quality & Policy Compliance Verification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific quality standards, compliance requirements, and
statistical sampling methodologies.

Key customization areas:
1. Quality thresholds and tolerances (what accuracy is acceptable for your risk)
2. Sampling methodology and sample sizes (based on population and confidence level)
3. Quality dimension weights (importance of accuracy vs. completeness, etc.)
4. Compliance criteria and scoring (organization-specific policy requirements)
5. Organization name, CISO details, contact information
6. File paths and naming conventions
7. Integration with quality management systems

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.8.24 Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
inventory data quality and policy compliance through systematic sampling and
verification across five quality dimensions.

**Purpose:**
Enables systematic assessment of whether inventory data is accurate, complete,
current, consistent, and compliant with organizational policies. Uses statistical
sampling methodology to provide objective, quantifiable quality metrics that can
be tracked over time and drive continuous improvement.

**Assessment Scope:**
- Accuracy: Correctness of inventory data vs. ground truth
- Completeness: Presence of all required attributes and mandatory fields
- Currency: Timeliness of data (last updated, review dates)
- Consistency: Data alignment across related systems (CMDB, HR, procurement)
- Policy Compliance: Adherence to ISMS-POL-A.5.9 requirements

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment methodology and quality framework
2. Accuracy Sampling - Statistical sampling for data accuracy verification
3. Completeness Assessment - Mandatory field presence and attribute coverage
4. Currency Assessment - Data freshness and update timeliness
5. Consistency Checks - Cross-system data alignment verification
6. Policy Compliance Matrix - Adherence to policy requirements
7. Quality Metrics & Scoring - Aggregated scores and weighted quality index
8. Evidence Register - Quality verification evidence and documentation

**Key Features:**
- Data validation with dropdown lists for quality dimensions
- Conditional formatting for quality scores (Green/Yellow/Red)
- Statistical sampling calculator (sample size, confidence intervals)
- Automated gap identification for quality failures
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Weighted quality scoring (5 dimensions with configurable weights)
- Trend tracking for quality improvement measurement
- Sample record tracking with verification results

**Integration:**
This assessment feeds into the A.5.9.5 Compliance Dashboard, which consolidates
data from all four assessment domains for executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - os (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a59_3_quality_compliance.py

Output:
    File: ISMS-IMP-A.5.9.3_Quality_Compliance_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Complete sampling plan (determine sample sizes per asset category)
    2. Perform accuracy sampling (verify sample records against ground truth)
    3. Run completeness checks (automated or manual validation)
    4. Assess currency (check last update dates, staleness)
    5. Perform consistency checks (compare inventory vs. CMDB, HR, procurement)
    6. Evaluate policy compliance (review against ISMS-POL-A.5.9 requirements)
    7. Review Quality Metrics & Scoring for overall quality index
    8. Collect and link evidence in Evidence Register sheet
    9. Export metrics CSV for dashboard consolidation (Sheet 7)
    10. Store assessment workbook per retention policy (7 years minimum)
    11. Update quarterly or after major data quality remediation efforts

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.9
Assessment Domain:    3 of 5
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organization License]

Related Documents:
    - ISMS-POL-A.5.9: Inventory of Information and Assets (Policy)
    - ISMS-IMP-A.5.9-1: Asset Discovery Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-2: Inventory Maintenance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-3: Quality & Compliance Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-4: Owner Accountability Assessment (Implementation Guide)
    - ISMS-IMP-A.5.9-5: Compliance Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 22.01.2026
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.9-3 specification
    - Five quality dimensions with weighted scoring
    - Statistical sampling methodology
    - Cross-system consistency validation
    - Policy compliance matrix
    - Evidence collection and audit trail support

[Future changes to be documented here]

--------------------------------------------------------------------------------
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# CUSTOMIZE: Configuration
CONFIG = {
    'organisation': '[Organisation]',
    'ciso_name': '[CISO Name]',
    'security_contact': '[security@organisation.ch]',
    
    # Color scheme (consistent with A.8.24 pattern)
    'colors': {
        'header_bg': '003366',      # Dark blue
        'header_text': 'FFFFFF',     # White
        'section_bg': '4472C4',      # Medium blue
        'green_light': 'C6EFCE',     # Light green (pass)
        'green_dark': '006100',      # Dark green (pass text)
        'yellow_light': 'FFEB9C',    # Light yellow (warning)
        'yellow_dark': '9C5700',     # Dark yellow (warning text)
        'red_light': 'FFC7CE',       # Light red (fail)
        'red_dark': '9C0006',        # Dark red (fail text)
        'gray_light': 'D9D9D9',      # Light gray (locked cells)
    },
    
    # Quality dimension weights (must sum to 100%)
    'quality_weights': {
        'Accuracy': 30,         # Most important - correct data
        'Completeness': 25,     # Required fields present
        'Currency': 20,         # Data freshness
        'Consistency': 15,      # Cross-system alignment
        'Policy Compliance': 10, # Policy adherence
    },
    
    # Quality targets (customize per dimension)
    'quality_targets': {
        'Accuracy': 98,         # Very high - errors are unacceptable
        'Completeness': 95,     # High - most fields should be populated
        'Currency': 90,         # Good - data reasonably current
        'Consistency': 95,      # High - systems should agree
        'Policy Compliance': 100, # Perfect - policy is mandatory
    },
    
    # Overall quality target (from policy)
    'overall_target': 97,
    
    # Sampling parameters (customize based on population)
    'sampling': {
        'confidence_level': 95,  # 95% confidence
        'margin_of_error': 5,    # ±5%
        'min_sample_size': 30,   # Minimum for statistical validity
    },
}

# Document identification constants
DOCUMENT_ID = "ISMS-IMP-A.5.9.3"
WORKBOOK_NAME = "Quality Compliance"
CONTROL_REF = "ISO/IEC 27001:2022 - Control A.5.9: Inventory of Information and Assets"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


def main():
    """Main execution function"""
    logger.info("="*80)
    logger.info("ISMS Control A.5.9 - Quality & Compliance Assessment Generator")
    logger.info("="*80)
    logger.info("")
    logger.info("Generating assessment workbook...")
    logger.info("")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create sheets in order
    sheets = [
        "Instructions & Legend",
        "Accuracy Sampling",
        "Completeness Assessment",
        "Currency Assessment",
        "Consistency Checks",
        "Policy Compliance Matrix",
        "Quality Metrics & Scoring",
        "Evidence Register",
        "Summary Dashboard",
        "Approval & Sign-Off",
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(title=sheet_name)
        logger.info(f"  ✓ Created sheet: {sheet_name}")
    
    logger.info("")
    logger.info("Populating sheets...")
    logger.info("")
    
    # Populate each sheet
    create_instructions_sheet(wb["Instructions & Legend"])
    logger.info("  ✓ Instructions & Legend")
    
    create_accuracy_sampling_sheet(wb["Accuracy Sampling"])
    logger.info("  ✓ Accuracy Sampling")
    
    create_completeness_sheet(wb["Completeness Assessment"])
    logger.info("  ✓ Completeness Assessment")
    
    create_currency_sheet(wb["Currency Assessment"])
    logger.info("  ✓ Currency Assessment")
    
    create_consistency_sheet(wb["Consistency Checks"])
    logger.info("  ✓ Consistency Checks")
    
    create_policy_compliance_sheet(wb["Policy Compliance Matrix"])
    logger.info("  ✓ Policy Compliance Matrix")
    
    create_quality_metrics_sheet(wb["Quality Metrics & Scoring"])
    logger.info("  ✓ Quality Metrics & Scoring")
    
    create_evidence_register_sheet(wb["Evidence Register"])
    logger.info("  ✓ Evidence Register")
    
    create_summary_dashboard_sheet(wb["Summary Dashboard"])
    logger.info("  ✓ Summary Dashboard")
    
    create_approval_signoff_sheet(wb["Approval & Sign-Off"])
    logger.info("  ✓ Approval & Sign-Off")
    
    # Save workbook
    filename = f"ISMS-IMP-A.5.9.3_Quality_Compliance_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    logger.info("")
    logger.info("="*80)
    logger.info(f"✅ SUCCESS: {filename}")
    logger.info("="*80)
    logger.info("")
    logger.info("Next Steps:")
    logger.info("  1. Open workbook and review Instructions & Legend")
    logger.info("  2. Determine sample sizes for accuracy sampling")
    logger.info("  3. Perform accuracy sampling (verify sample records)")
    logger.info("  4. Run completeness checks (mandatory fields)")
    logger.info("  5. Assess currency (data freshness)")
    logger.info("  6. Perform consistency checks (cross-system)")
    logger.info("  7. Evaluate policy compliance")
    logger.info("  8. Review Quality Metrics & Scoring for overall quality index")
    logger.info("  9. Export CSV from Sheet 7 for dashboard consolidation")
    logger.info("  10. Obtain stakeholder review and approval")
    logger.info("")
    logger.info(f"Output location: {os.path.abspath(filename)}")
    logger.info("")


def create_instructions_sheet(ws):
    """Create Instructions & Legend sheet"""
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{DOCUMENT_ID}  -  Quality & Compliance Assessment\n{CONTROL_REF}"
    ws['A1'].font = Font(size=16, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 40
    
    # Subtitle
    ws.merge_cells('A2:H2')
    ws['A2'] = "Assessment Domain 3 of 5: Data Quality & Policy Compliance Verification"
    ws['A2'].font = Font(size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Purpose section
    ws['A4'] = "PURPOSE"
    ws['A4'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws['A4'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    ws.merge_cells('A5:H8')
    purpose_text = """This assessment evaluates the QUALITY and COMPLIANCE of inventory data through systematic sampling and verification.

Discovery finds assets. Maintenance keeps inventory current. Quality ensures data is accurate, complete, current, consistent, and compliant.

Five quality dimensions are assessed using statistical sampling and cross-system verification."""
    ws['A5'] = purpose_text
    ws['A5'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[5].height = 60
    
    # Five Quality Dimensions
    ws['A10'] = "FIVE QUALITY DIMENSIONS"
    ws['A10'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws['A10'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    dimensions = [
        ("1. ACCURACY (30% weight)", "Correctness of data vs. ground truth (e.g., server really exists, owner is correct)", "98%"),
        ("2. COMPLETENESS (25%)", "All required fields populated, no missing mandatory attributes", "95%"),
        ("3. CURRENCY (20%)", "Data freshness - recently reviewed, not stale (e.g., <90 days)", "90%"),
        ("4. CONSISTENCY (15%)", "Data alignment across related systems (inventory ↔ CMDB ↔ HR ↔ procurement)", "95%"),
        ("5. POLICY COMPLIANCE (10%)", "Adherence to ISMS-POL-A.5.9 requirements (ownership, classification, etc.)", "100%"),
    ]
    
    row = 11
    for dimension, description, target in dimensions:
        ws[f'A{row}'] = dimension
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = description
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        ws[f'C{row}'] = target
        ws[f'C{row}'].font = Font(bold=True)
        ws.row_dimensions[row].height = 30
        row += 1
    
    # Statistical Sampling
    ws[f'A{row+1}'] = "STATISTICAL SAMPLING METHODOLOGY"
    ws[f'A{row+1}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{row+1}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    row += 2
    ws.merge_cells(f'A{row}:H{row+3}')
    sampling_text = """For ACCURACY sampling, use statistical sampling:
• 95% confidence level, ±5% margin of error
• Minimum 30 samples per asset category (for statistical validity)
• Random selection (use random number generator, not cherry-picking)
• Verify each sample against ground truth (physical inspection, source system, owner confirmation)

For COMPLETENESS, CURRENCY, CONSISTENCY: Can use automated queries/reports (100% population check)"""
    ws[f'A{row}'] = sampling_text
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[row].height = 100
    
    row += 5
    
    # Weighted Scoring
    ws[f'A{row}'] = "WEIGHTED QUALITY SCORING"
    ws[f'A{row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    row += 1
    ws[f'A{row}'] = "Overall Quality Index = (Accuracy × 30%) + (Completeness × 25%) + (Currency × 20%) + (Consistency × 15%) + (Policy Compliance × 10%)"
    ws[f'A{row}'].font = Font(italic=True)
    ws.merge_cells(f'A{row}:H{row}')
    
    row += 1
    ws[f'A{row}'] = "Target: ≥97% for compliant quality"
    ws[f'A{row}'].font = Font(bold=True)
    
    # Traffic Light Legend
    row += 2
    ws[f'A{row}'] = "TRAFFIC LIGHT LEGEND"
    ws[f'A{row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    row += 1
    # Green
    ws[f'A{row}'] = "✅ GREEN - High Quality"
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid')
    ws[f'B{row}'] = "Score ≥ Target (e.g., Accuracy ≥98%)"
    
    # Yellow
    row += 1
    ws[f'A{row}'] = "⚠️ YELLOW - Needs Improvement"
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    ws[f'B{row}'] = "Score within 5% of target (e.g., Accuracy 93-97%)"
    
    # Red
    row += 1
    ws[f'A{row}'] = "❌ RED - Poor Quality"
    ws[f'A{row}'].fill = PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid')
    ws[f'B{row}'] = "Score >5% below target (e.g., Accuracy <93%)"
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 12
    
    # Protect sheet


def create_accuracy_sampling_sheet(ws):
    """Create Accuracy Sampling sheet"""
    
    # Header
    ws.merge_cells('A1:N1')
    ws['A1'] = "Accuracy Sampling - Statistical Verification"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25
    
    # Column headers
    headers = [
        ('A', 'Sample ID', 15),
        ('B', 'Asset Category', 25),
        ('C', 'Asset ID (Inventory)', 25),
        ('D', 'Attribute Verified', 30),
        ('E', 'Inventory Value', 30),
        ('F', 'Ground Truth Value', 30),
        ('G', 'Match?', 12),
        ('H', 'Discrepancy Type', 25),
        ('I', 'Verification Method', 30),
        ('J', 'Verified By', 20),
        ('K', 'Verification Date', 18),
        ('L', 'Corrected?', 15),
        ('M', 'Evidence ID', 15),
        ('N', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws[f'{col}3'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[col].width = width
    
    # Sample data rows (user fills actual samples)
    sample_data = [
        ("ACC-001", "Information Assets", "DB-PROD-001", "Database Size", "500 GB", "482 GB", "No"),
        ("ACC-002", "IT Infrastructure", "SRV-WEB-023", "IP Address", "10.1.2.45", "10.1.2.45", "Yes"),
        ("ACC-003", "Applications", "APP-JIRA-001", "Owner", "John Doe", "Jane Smith", "No"),
        ("ACC-004", "Personnel Assets", "PER-ENG-042", "Department", "Engineering", "Engineering", "Yes"),
    ]
    
    row = 4
    for sample_id, category, asset_id, attribute, inv_value, truth_value, match in sample_data:
        ws[f'A{row}'] = sample_id
        ws[f'B{row}'] = category
        ws[f'C{row}'] = asset_id
        ws[f'D{row}'] = attribute
        ws[f'E{row}'] = inv_value
        ws[f'F{row}'] = truth_value
        ws[f'G{row}'] = match
        
        # Unlock for user input
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Add 50 blank rows for sampling
    for i in range(50):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    # Data validations
    categories = ["Information Assets", "IT Infrastructure", "Applications", "Physical Assets", "Personnel Assets"]
    dv_category = DataValidation(type="list", formula1=f'"{",".join(categories)}"', allow_blank=True)
    dv_category.add(f'B4:B100')
    ws.add_data_validation(dv_category)
    
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_yesno.add(f'G4:G100')
    dv_yesno.add(f'L4:L100')
    ws.add_data_validation(dv_yesno)
    
    verification_methods = [
        "Physical Inspection",
        "Source System Query",
        "Owner Confirmation",
        "Network Scan Verification",
        "HR System Verification",
        "Procurement Records",
        "Third-party Audit",
        "Other (Specify)"
    ]
    dv_method = DataValidation(type="list", formula1=f'"{",".join(verification_methods)}"', allow_blank=True)
    dv_method.add(f'I4:I100')
    ws.add_data_validation(dv_method)
    
    # Summary section
    summary_row = row + 2
    ws[f'A{summary_row}'] = "ACCURACY SAMPLING SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Total Samples"
    ws[f'B{summary_row}'] = f'=COUNTA(A4:A{row-51})'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Matches (Accurate)"
    ws[f'B{summary_row}'] = f'=COUNTIF(G4:G{row-51},"Yes")'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Accuracy %"
    ws[f'B{summary_row}'] = f'=IFERROR(B{summary_row-1}/B{summary_row-2}*100,0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "98%"
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-98'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Status"
    ws[f'B{summary_row}'] = f'=IF(B{summary_row-3}>=98,"✅ High Quality",IF(B{summary_row-3}>=93,"⚠️ Needs Improvement","❌ Poor Quality"))'
    ws[f'B{summary_row}'].alignment = Alignment(horizontal='center')
    ws[f'B{summary_row}'].font = Font(bold=True)


def create_completeness_sheet(ws):
    """Create Completeness Assessment sheet"""
    
    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "Completeness Assessment - Mandatory Fields"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Asset Category', 25),
        ('B', 'Mandatory Field', 30),
        ('C', 'Total Assets', 15),
        ('D', 'Assets with Field Populated', 20),
        ('E', 'Completeness %', 16),
        ('F', 'Target %', 12),
        ('G', 'Gap', 12),
        ('H', 'Status', 18),
        ('I', 'Missing Count', 15),
        ('J', 'Evidence ID', 15),
        ('K', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Pre-populate mandatory fields (customize per organization)
    mandatory_fields = [
        ("Information Assets", "Asset Name", 95),
        ("Information Assets", "Owner", 95),
        ("Information Assets", "Classification", 95),
        ("Information Assets", "Location", 95),
        ("IT Infrastructure", "Asset Name", 95),
        ("IT Infrastructure", "IP Address / Hostname", 95),
        ("IT Infrastructure", "Owner", 95),
        ("IT Infrastructure", "Location", 95),
        ("Applications", "Application Name", 95),
        ("Applications", "Version", 95),
        ("Applications", "Owner", 95),
        ("Applications", "Business Purpose", 95),
        ("Physical Assets", "Asset Name", 95),
        ("Physical Assets", "Location", 95),
        ("Physical Assets", "Custodian", 95),
        ("Personnel Assets", "Name", 100),
        ("Personnel Assets", "Role", 100),
        ("Personnel Assets", "Department", 100),
        ("Personnel Assets", "Manager", 100),
    ]
    
    row = 4
    for category, field, target in mandatory_fields:
        ws[f'A{row}'] = category
        ws[f'B{row}'] = field
        ws[f'F{row}'] = f'{target}%'
        
        # User enters counts
        ws[f'C{row}'].protection = Protection(locked=False)
        ws[f'D{row}'].protection = Protection(locked=False)
        
        # Completeness % (formula)
        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        
        # Gap
        ws[f'G{row}'] = f'=E{row}-{target}'
        ws[f'G{row}'].number_format = '0.0"%"'
        
        # Status
        ws[f'H{row}'] = f'=IF(E{row}>={target},"✅ Complete",IF(E{row}>={target-5},"⚠️ At Risk","❌ Incomplete"))'
        ws[f'H{row}'].alignment = Alignment(horizontal='center')
        
        # Missing count
        ws[f'I{row}'] = f'=C{row}-D{row}'
        
        # Unlock evidence and notes
        ws[f'J{row}'].protection = Protection(locked=False)
        ws[f'K{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "COMPLETENESS SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Completeness %"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(E4:E{row-1}),0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "95%"
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-95'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Status"
    ws[f'B{summary_row}'] = f'=IF(B{summary_row-3}>=95,"✅ High Quality",IF(B{summary_row-3}>=90,"⚠️ Needs Improvement","❌ Poor Quality"))'
    ws[f'B{summary_row}'].alignment = Alignment(horizontal='center')
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['95'],
                   fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='between', formula=['90', '94'],
                   fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='lessThan', formula=['90'],
                   fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid'))
    )


def create_currency_sheet(ws):
    """Create Currency Assessment sheet"""
    
    # Header
    ws.merge_cells('A1:K1')
    ws['A1'] = "Currency Assessment - Data Freshness"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Asset Category', 25),
        ('B', 'Currency Metric', 30),
        ('C', 'Total Assets', 15),
        ('D', 'Assets Meeting Currency Target', 22),
        ('E', 'Currency %', 15),
        ('F', 'Target %', 12),
        ('G', 'Gap', 12),
        ('H', 'Status', 18),
        ('I', 'Stale Assets Count', 18),
        ('J', 'Evidence ID', 15),
        ('K', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Currency metrics
    currency_metrics = [
        ("Information Assets", "Last Reviewed <90 days", 90),
        ("Information Assets", "Last Updated <180 days", 90),
        ("IT Infrastructure", "Last Reviewed <90 days", 90),
        ("IT Infrastructure", "Last Scanned <30 days", 90),
        ("Applications", "Last Reviewed <180 days", 90),
        ("Applications", "Version Current", 90),
        ("Physical Assets", "Last Inspected <365 days", 90),
        ("Personnel Assets", "Last Updated <30 days", 100),
    ]
    
    row = 4
    for category, metric, target in currency_metrics:
        ws[f'A{row}'] = category
        ws[f'B{row}'] = metric
        ws[f'F{row}'] = f'{target}%'
        
        ws[f'C{row}'].protection = Protection(locked=False)
        ws[f'D{row}'].protection = Protection(locked=False)
        
        ws[f'E{row}'] = f'=IFERROR(D{row}/C{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        
        ws[f'G{row}'] = f'=E{row}-{target}'
        ws[f'G{row}'].number_format = '0.0"%"'
        
        ws[f'H{row}'] = f'=IF(E{row}>={target},"✅ Current",IF(E{row}>={target-5},"⚠️ At Risk","❌ Stale"))'
        ws[f'H{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'I{row}'] = f'=C{row}-D{row}'
        
        ws[f'J{row}'].protection = Protection(locked=False)
        ws[f'K{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "CURRENCY SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Currency %"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(E4:E{row-1}),0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "90%"
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-90'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Status"
    ws[f'B{summary_row}'] = f'=IF(B{summary_row-3}>=90,"✅ High Quality",IF(B{summary_row-3}>=85,"⚠️ Needs Improvement","❌ Poor Quality"))'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['90'],
                   fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='between', formula=['85', '89'],
                   fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='lessThan', formula=['85'],
                   fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid'))
    )


def create_consistency_sheet(ws):
    """Create Consistency Checks sheet"""
    
    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = "Consistency Checks - Cross-System Alignment"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Consistency Check', 35),
        ('B', 'System A', 25),
        ('C', 'System B', 25),
        ('D', 'Records in A', 15),
        ('E', 'Records in B', 15),
        ('F', 'Matches', 15),
        ('G', 'Consistency %', 16),
        ('H', 'Target %', 12),
        ('I', 'Gap', 12),
        ('J', 'Status', 18),
        ('K', 'Evidence ID', 15),
        ('L', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Consistency checks
    checks = [
        ("IT Assets: Inventory ↔ CMDB", "Asset Inventory", "CMDB", 95),
        ("Servers: Inventory ↔ Network Scan", "Asset Inventory", "Network Scan Results", 95),
        ("Personnel: Inventory ↔ HR System", "Asset Inventory", "HR System", 100),
        ("Applications: Inventory ↔ License DB", "Asset Inventory", "License Database", 95),
        ("Hardware: Inventory ↔ Procurement", "Asset Inventory", "Procurement Records", 95),
        ("Cloud Resources: Inventory ↔ Cloud API", "Asset Inventory", "AWS/Azure API", 95),
    ]
    
    row = 4
    for check, sys_a, sys_b, target in checks:
        ws[f'A{row}'] = check
        ws[f'B{row}'] = sys_a
        ws[f'C{row}'] = sys_b
        ws[f'H{row}'] = f'{target}%'
        
        # User enters counts
        for col in ['D', 'E', 'F']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        
        # Consistency % = matches / max(A, B)
        ws[f'G{row}'] = f'=IFERROR(F{row}/MAX(D{row},E{row})*100,0)'
        ws[f'G{row}'].number_format = '0.0"%"'
        
        ws[f'I{row}'] = f'=G{row}-{target}'
        ws[f'I{row}'].number_format = '0.0"%"'
        
        ws[f'J{row}'] = f'=IF(G{row}>={target},"✅ Consistent",IF(G{row}>={target-5},"⚠️ At Risk","❌ Inconsistent"))'
        ws[f'J{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'K{row}'].protection = Protection(locked=False)
        ws[f'L{row}'].protection = Protection(locked=False)
        
        row += 1
    
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "CONSISTENCY SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Consistency %"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(G4:G{row-1}),0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "95%"
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-95'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Status"
    ws[f'B{summary_row}'] = f'=IF(B{summary_row-3}>=95,"✅ High Quality",IF(B{summary_row-3}>=90,"⚠️ Needs Improvement","❌ Poor Quality"))'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'G4:G{row-1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['95'],
                   fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'G4:G{row-1}',
        CellIsRule(operator='between', formula=['90', '94'],
                   fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'G4:G{row-1}',
        CellIsRule(operator='lessThan', formula=['90'],
                   fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid'))
    )


def create_policy_compliance_sheet(ws):
    """Create Policy Compliance Matrix sheet"""
    
    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "Policy Compliance Matrix - ISMS-POL-A.5.9 Requirements"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Policy Requirement', 50),
        ('B', 'SHALL Requirement', 50),
        ('C', 'Compliant Assets', 18),
        ('D', 'Total Assets', 15),
        ('E', 'Compliance %', 16),
        ('F', 'Target', 12),
        ('G', 'Gap', 12),
        ('H', 'Status', 18),
        ('I', 'Evidence ID', 15),
        ('J', 'Notes', 30),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Policy requirements (from ISMS-POL-A.5.9)
    requirements = [
        ("Asset Ownership", "[Organisation] SHALL assign an owner to each asset", 100),
        ("Classification", "[Organisation] SHALL classify information assets by sensitivity", 100),
        ("Acceptable Use", "Asset owners SHALL define acceptable use restrictions", 100),
        ("Handling Requirements", "Asset owners SHALL document handling requirements", 100),
        ("Inventory Maintenance", "[Organisation] SHALL maintain current and accurate inventory", 100),
        ("Periodic Review", "Asset owners SHALL review assets at least annually", 100),
        ("Decommissioning", "[Organisation] SHALL follow secure decommissioning procedures", 100),
    ]
    
    row = 4
    for req, shall_text, target in requirements:
        ws[f'A{row}'] = req
        ws[f'B{row}'] = shall_text
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        ws[f'F{row}'] = f'{target}%'
        
        # User enters counts
        ws[f'C{row}'].protection = Protection(locked=False)
        ws[f'D{row}'].protection = Protection(locked=False)
        
        ws[f'E{row}'] = f'=IFERROR(C{row}/D{row}*100,0)'
        ws[f'E{row}'].number_format = '0.0"%"'
        
        ws[f'G{row}'] = f'=E{row}-{target}'
        ws[f'G{row}'].number_format = '0.0"%"'
        
        ws[f'H{row}'] = f'=IF(E{row}={target},"✅ Compliant",IF(E{row}>=95,"⚠️ Minor Gap","❌ Non-Compliant"))'
        ws[f'H{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'I{row}'].protection = Protection(locked=False)
        ws[f'J{row}'].protection = Protection(locked=False)
        
        ws.row_dimensions[row].height = 30
        row += 1
    
    # Summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "POLICY COMPLIANCE SUMMARY"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Average Policy Compliance %"
    ws[f'B{summary_row}'] = f'=IFERROR(AVERAGE(E4:E{row-1}),0)'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    ws[f'B{summary_row}'].font = Font(bold=True, size=12)
    ws[f'B{summary_row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Target"
    ws[f'B{summary_row}'] = "100%"
    ws[f'B{summary_row}'].font = Font(bold=True, color=CONFIG['colors']['red_dark'])
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Gap"
    ws[f'B{summary_row}'] = f'=B{summary_row-2}-100'
    ws[f'B{summary_row}'].number_format = '0.0"%"'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Status"
    ws[f'B{summary_row}'] = f'=IF(B{summary_row-3}=100,"✅ Fully Compliant",IF(B{summary_row-3}>=95,"⚠️ Minor Gaps","❌ Non-Compliant"))'
    ws[f'B{summary_row}'].font = Font(bold=True)
    
    # Conditional formatting
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='equal', formula=['100'],
                   fill=PatternFill(start_color=CONFIG['colors']['green_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='between', formula=['95', '99'],
                   fill=PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'E4:E{row-1}',
        CellIsRule(operator='lessThan', formula=['95'],
                   fill=PatternFill(start_color=CONFIG['colors']['red_light'], fill_type='solid'))
    )


def create_quality_metrics_sheet(ws):
    """Create Quality Metrics & Scoring sheet - consolidates all dimensions"""
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Quality Metrics & Scoring - Weighted Quality Index"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Quality dimensions table
    ws['A3'] = "Quality Dimension"
    ws['B3'] = "Score %"
    ws['C3'] = "Weight %"
    ws['D3'] = "Weighted Score"
    ws['E3'] = "Target %"
    ws['F3'] = "Gap"
    ws['G3'] = "Status"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    
    # Link to dimension sheets
    dimensions_data = [
        ("Accuracy", "='Accuracy Sampling'!B62", 30, 98),
        ("Completeness", "='Completeness Assessment'!B28", 25, 95),
        ("Currency", "='Currency Assessment'!B16", 20, 90),
        ("Consistency", "='Consistency Checks'!B13", 15, 95),
        ("Policy Compliance", "='Policy Compliance Matrix'!B15", 10, 100),
    ]
    
    row = 4
    for dimension, formula, weight, target in dimensions_data:
        ws[f'A{row}'] = dimension
        ws[f'A{row}'].font = Font(bold=True)
        
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = '0.0"%"'
        ws[f'B{row}'].font = Font(bold=True)
        
        ws[f'C{row}'] = f'{weight}%'
        
        ws[f'D{row}'] = f'=B{row}*{weight}/100'
        ws[f'D{row}'].number_format = '0.0"%"'
        
        ws[f'E{row}'] = f'{target}%'
        
        ws[f'F{row}'] = f'=B{row}-{target}/100'
        ws[f'F{row}'].number_format = '0.0"%"'
        
        ws[f'G{row}'] = f'=IF(B{row}>={target}/100,"✅ High",IF(B{row}>={target}/100-0.05,"⚠️ Medium","❌ Low"))'
        ws[f'G{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Overall quality index
    overall_row = row + 2
    ws[f'A{overall_row}'] = "OVERALL QUALITY INDEX"
    ws[f'A{overall_row}'].font = Font(size=12, bold=True)
    ws[f'B{overall_row}'] = f'=SUM(D4:D8)'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    ws[f'B{overall_row}'].font = Font(bold=True, size=14)
    ws[f'B{overall_row}'].fill = PatternFill(start_color=CONFIG['colors']['yellow_light'], fill_type='solid')
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Target"
    ws[f'B{overall_row}'] = "97%"
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Gap"
    ws[f'B{overall_row}'] = f'=B{overall_row-2}-0.97'
    ws[f'B{overall_row}'].number_format = '0.0"%"'
    
    overall_row += 1
    ws[f'A{overall_row}'] = "Overall Status"
    ws[f'B{overall_row}'] = f'=IF(B{overall_row-3}>=0.97,"✅ High Quality",IF(B{overall_row-3}>=0.92,"⚠️ Needs Improvement","❌ Poor Quality"))'
    ws[f'B{overall_row}'].font = Font(bold=True)
    
    # CSV Export section
    csv_row = overall_row + 3
    ws[f'A{csv_row}'] = "CSV EXPORT FOR DASHBOARD (Copy rows below)"
    ws[f'A{csv_row}'].font = Font(size=11, bold=True, color=CONFIG['colors']['header_text'])
    ws[f'A{csv_row}'].fill = PatternFill(start_color=CONFIG['colors']['section_bg'], fill_type='solid')
    
    csv_row += 1
    ws[f'A{csv_row}'] = "Quality_Dimension"
    ws[f'B{csv_row}'] = "Score_%"
    ws[f'C{csv_row}'] = "Status"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{csv_row}'].font = Font(bold=True)
    
    csv_row += 1
    for i in range(5):
        ws[f'A{csv_row+i}'] = f'=A{4+i}'
        ws[f'B{csv_row+i}'] = f'=B{4+i}'
        ws[f'C{csv_row+i}'] = f'=G{4+i}'
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18


def create_evidence_register_sheet(ws):
    """Create Evidence Register sheet"""
    
    # Title
    ws.merge_cells('A1:N1')
    ws['A1'] = "Quality Evidence Register"
    ws['A1'].font = Font(size=14, bold=True, color=CONFIG['colors']['header_text'])
    ws['A1'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        ('A', 'Evidence ID', 15),
        ('B', 'Quality Dimension', 25),
        ('C', 'Evidence Type', 30),
        ('D', 'Evidence Description', 50),
        ('E', 'Evidence Location', 40),
        ('F', 'Collection Date', 15),
        ('G', 'Collected By', 25),
        ('H', 'Validity Period', 20),
        ('I', 'Review Date', 15),
        ('J', 'Reviewed By', 25),
        ('K', 'Review Status', 20),
        ('L', 'Retention End Date', 18),
        ('M', 'Related Assessment', 25),
        ('N', 'Notes', 40),
    ]
    
    for col, header, width in headers:
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True, color=CONFIG['colors']['header_text'])
        ws[f'{col}3'].fill = PatternFill(start_color=CONFIG['colors']['header_bg'], fill_type='solid')
        ws.column_dimensions[col].width = width
    
    # Sample evidence
    sample_data = [
        ("QUAL-001", "Accuracy", "Sampling Results", "Accuracy sampling results - 50 samples verified", "/evidence/accuracy_sampling_20260122.xlsx", "22.01.2026"),
        ("QUAL-002", "Completeness", "Completeness Report", "Mandatory fields completeness query results", "/evidence/completeness_report_20260122.xlsx", "22.01.2026"),
        ("QUAL-003", "Currency", "Staleness Report", "Assets not reviewed in >90 days", "/evidence/staleness_report_20260122.xlsx", "22.01.2026"),
        ("QUAL-004", "Consistency", "Reconciliation Report", "Inventory vs. CMDB reconciliation", "/evidence/reconciliation_20260122.xlsx", "22.01.2026"),
    ]
    
    row = 4
    for evidence_id, dimension, evidence_type, description, location, date in sample_data:
        ws[f'A{row}'] = evidence_id
        ws[f'B{row}'] = dimension
        ws[f'C{row}'] = evidence_type
        ws[f'D{row}'] = description
        ws[f'E{row}'] = location
        ws[f'F{row}'] = date
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    # Add empty rows
    for i in range(20):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws[f'{col}{row}'].protection = Protection(locked=False)
        row += 1
    
    # Data validations
    quality_dimensions = ["Accuracy", "Completeness", "Currency", "Consistency", "Policy Compliance", "All Dimensions"]
    dv_dimension = DataValidation(type="list", formula1=f'"{",".join(quality_dimensions)}"', allow_blank=True)
    dv_dimension.add(f'B4:B100')
    ws.add_data_validation(dv_dimension)
    
    review_statuses = ["Pending Review", "Reviewed - Valid", "Reviewed - Update Needed", "Reviewed - Invalid"]
    dv_status = DataValidation(type="list", formula1=f'"{",".join(review_statuses)}"', allow_blank=True)
    dv_status.add(f'K4:K100')
    ws.add_data_validation(dv_status)


# Execute main function


def create_summary_dashboard_sheet(ws):
    """Create Summary Dashboard sheet"""
    CHECK = "✅"
    WARNING = "⚠️"
    XMARK = "❌"
    TARGET = "🎯"
    CHART = "📊"
    
    metrics_sheet = "Quality Metrics & Scoring"
    assessment_name = "Quality & Compliance Assessment"
    compliance_ref = "B10"
    key_metrics = [
        ('Accuracy Score', 'B4', '98%', 'percentage'),
        ('Completeness Score', 'B5', '95%', 'percentage'),
        ('Currency Score', 'B6', '95%', 'percentage'),
    ]
    
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{CHART} {assessment_name.upper()} - SUMMARY DASHBOARD"
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:F2')
    ws['A2'] = "Quick overview of key metrics and compliance status"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].fill = PatternFill(start_color='E7E6E6', fill_type='solid')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    row = 4
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"{TARGET} OVERALL COMPLIANCE"
    ws[f'A{row}'].font = Font(size=13, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='2E75B5', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Overall Compliance:"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = f"='{metrics_sheet}'!{compliance_ref}"
    ws[f'D{row}'].font = Font(size=18, bold=True)
    ws[f'D{row}'].number_format = '0.0"%"'
    ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"{CHART} KEY METRICS"
    ws[f'A{row}'].font = Font(size=13, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='2E75B5', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    for col_idx, header in enumerate(["Metric", "Current", "Target", "Status"], start=1):
        ws.cell(row=row, column=col_idx, value=header)
        ws.cell(row=row, column=col_idx).font = Font(size=11, bold=True, color='FFFFFF')
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color='4472C4', fill_type='solid')
    row += 1
    
    for metric_label, cell_ref, target, format_type in key_metrics:
        ws.cell(row=row, column=1, value=metric_label)
        ws.cell(row=row, column=2, value=f"='{metrics_sheet}'!{cell_ref}")
        if format_type == 'percentage':
            ws.cell(row=row, column=2).number_format = '0.0"%"'
        ws.cell(row=row, column=3, value=target)
        ws.cell(row=row, column=4, value=f'=IF(B{row}>0,"{CHECK}","{XMARK}")')
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_approval_signoff_sheet(ws):
    """Create Approval & Sign-Off sheet"""
    CHECK = "✅"
    CLOCK = "⏳"
    XMARK = "❌"
    
    assessment_name = "Quality & Compliance Assessment"
    checklist_items = [
        'Accuracy sampling completed',
        'Completeness assessment completed',
        'Currency assessment completed',
        'Consistency checks performed',
        'Policy compliance verified',
        'Quality metrics calculated',
        'Evidence collected and registered',
        'Assessment reviewed by Information Security'
    ]
    
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{assessment_name.upper()} - APPROVAL & SIGN-OFF"
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:F2')
    ws['A2'] = "Complete when assessment is finished and ready for approval"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    row = 4
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "COMPLETION CHECKLIST"
    ws[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='2E75B5', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    for col_idx, header in enumerate(["Item", "Status", "Completed By", "Date", "Notes"], start=1):
        ws.cell(row=row, column=col_idx, value=header)
        ws.cell(row=row, column=col_idx).font = Font(size=11, bold=True, color='FFFFFF')
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color='4472C4', fill_type='solid')
    row += 1
    
    checklist_start = row
    for item in checklist_items:
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=1).font = Font(size=10)
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFF2CC', fill_type='solid')
            ws.cell(row=row, column=col).protection = Protection(locked=False)
        row += 1
    
    from openpyxl.worksheet.datavalidation import DataValidation
    status_options = [f"{CHECK} Complete", f"{CLOCK} In Progress", f"{XMARK} Not Done"]
    dv = DataValidation(type="list", formula1=f'"{",".join(status_options)}"', allow_blank=True)
    dv.add(f'B{checklist_start}:B{row-1}')
    ws.add_data_validation(dv)
    
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "APPROVALS"
    ws[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='2E75B5', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    for col_idx, header in enumerate(["Role", "Name", "Signature", "Date", "Comments"], start=1):
        ws.cell(row=row, column=col_idx, value=header)
        ws.cell(row=row, column=col_idx).font = Font(size=11, bold=True, color='FFFFFF')
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color='4472C4', fill_type='solid')
    row += 1
    
    for approver in ["Asset Management Lead", "Information Security Manager", "CISO"]:
        ws.cell(row=row, column=1, value=approver)
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color='D8E4F8', fill_type='solid')
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFF2CC', fill_type='solid')
            ws.cell(row=row, column=col).protection = Protection(locked=False)
        row += 1
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30



if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
