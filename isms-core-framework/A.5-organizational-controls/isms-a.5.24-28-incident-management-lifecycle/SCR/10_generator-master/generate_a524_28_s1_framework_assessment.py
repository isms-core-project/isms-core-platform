#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.24-28.S1 - Incident Management Framework Assessment Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.24: Incident Management Planning & Preparation
Assessment Domain 1 of 5: Framework & Governance

--------------------------------------------------------------------------------
PURPOSE & DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
the incident management framework and governance structure, focusing on the
planning and preparation phase (A.5.24) of the incident management lifecycle.

**Assessment Scope:**
- Governance framework (policy, procedures, authority)
- CSIRT/SOC organizational structure and staffing
- Role definitions, responsibilities, and RACI clarity
- Training and competency programs
- Tools and technology capabilities
- Integration with monitoring, logging, BC/DR, and other controls

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and methodology
2. Governance Assessment - 25 questions on policy, procedures, compliance
3. Organizational Structure - 30 questions on CSIRT model, roles, staffing
4. Training & Competency - 25 questions on programs, exercises, certifications
5. Tools & Technology - 30 questions on ticketing, SIEM, forensics, automation
6. Integration Assessment - 25 questions on cross-control integration
7. Gap Analysis - Gap prioritization and remediation planning
8. Evidence Register - Audit evidence tracking (100 items capacity)
9. Dashboard - Executive summary with maturity scoring
10. Approval Sign-Off - Multi-stakeholder approval workflow

**Key Features:**
- 117 assessment questions across 5 domains
- Data validation with dropdown lists
- Conditional formatting for gap identification
- Automated maturity scoring (Levels 1-5)
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This is Domain 1 of 5 assessment domains for A.5.24-28 Incident Management:
- S1: Framework & Governance (this assessment)
- S2: Detection & Classification (A.5.25 focus)
- S3: Response Capabilities (A.5.26 focus)
- S4: Forensic Evidence Management (A.5.28 focus)
- S5: Learning & Improvement (A.5.27 focus)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a524_28_s1_framework_assessment.py

Output:
    File: ISMS-IMP-A.5.24-28.S1_Framework_Assessment_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Review and customize assessment questions for your organization
    2. Complete document information in Instructions & Legend
    3. Conduct stakeholder interviews (CSIRT, Legal, HR, IT Ops)
    4. Fill yellow cells in each assessment sheet (Governance through Integration)
    5. Collect and link audit evidence (policy docs, org charts, training records)
    6. Review Gap Analysis and prioritize remediation
    7. Complete Dashboard review with management
    8. Obtain stakeholder approvals (CSIRT Manager, CISO)

Estimated Completion Time:
    - Information gathering: 3-4 hours
    - Assessment completion: 3-4 hours
    - Evidence collection: 1-2 hours
    - Quality review: 1-2 hours
    Total: 8-12 hours (depending on organization size/complexity)

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.5.24-28
Assessment Domain:    1 of 5 (Framework & Governance - A.5.24 Focus)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              Internal Use

Related Documents:
    - ISMS-POL-A.5.24-28: Incident Management Lifecycle Policy
    - ISMS-REF-A.5.24-28: Incident Response Reference Guide
    - ISMS-IMP-A.5.24-28.S2: Detection & Classification Assessment
    - ISMS-IMP-A.5.24-28.S3: Response Capabilities Assessment
    - ISMS-IMP-A.5.24-28.S4: Forensic Evidence Assessment
    - ISMS-IMP-A.5.24-28.S5: Learning & Improvement Assessment

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 2026-01-30
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.24-28.S1 specification
    - 117 assessment questions across 5 domains
    - Automated maturity scoring and gap analysis
    - Integration with evidence tracking

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Customization Required:**
This script generates a template assessment. Customize the following:
- Question relevance to your organizational context
- Dropdown options to match your CSIRT model and tools
- Evidence types to align with your audit requirements
- Maturity scoring thresholds based on your risk appetite

**Collaboration Essential:**
This assessment CANNOT be completed in isolation. Required stakeholders:
- CSIRT/SOC Manager (primary assessor)
- CISO (governance, budget)
- HR (staffing, training records)
- IT Operations (tools, integration)
- Legal/Compliance (regulatory requirements)

**Assessment Frequency:**
- Annual review (minimum)
- After major organizational changes (merger, restructuring)
- After major incidents (lessons learned)
- After audit findings

**Data Sensitivity:**
Assessment workbooks contain sensitive information:
- Organizational structure and staffing details
- Tool inventory and capabilities
- Security gaps and vulnerabilities
- Training records and competency data

Handle in accordance with your organization's data classification policies.

================================================================================
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)




# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.24-28.S1"
WORKBOOK_NAME = "Incident Management Framework"
CONTROL_ID = "A.5.24-28"
CONTROL_NAME = "Incident Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching specification."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure per markdown specification
    sheets = [
        "Instructions & Legend",
        "Governance Assessment",
        "Organizational Structure",
        "Training & Competency",
        "Tools & Technology",
        "Integration Assessment",
        "Gap Analysis",
        "Evidence Register",
        "Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color="D8E4F8", end_color="D8E4F8", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "calculated_cell": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
        "gap_critical": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "gap_high": {"fill": PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")},
        "gap_medium": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "gap_low": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
    }
    return styles


# ============================================================================
# SECTION 2: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create the Instructions & Legend sheet."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.5.24-28.S1 — Incident Management Framework Assessment\n"
        "ISO/IEC 27001:2022 - Control A.5.24: Planning and Preparation"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # Document Information
    row = 3
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:B{row}")

    doc_info = [
        ("Document ID", "ISMS-IMP-A.5.24-28.S1"),
        ("Assessment Area", "Incident Management Framework & Governance (Domain 1)"),
        ("Related Policy", "ISMS-POL-A.5.24-28, Section 2.1"),
        ("Version", "1.0"),
        ("Assessment Date", ""),  # User input
        ("Completed By", ""),     # User input
        ("Organization", ""),      # User input
        ("Review Cycle", "Annual"),
        ("Next Review Date", ""),  # Auto-calculate
    ]

    row += 1
    for label, value in doc_info:
        ws[f"A{row}"] = label + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if label in ["Assessment Date", "Completed By", "Organization"]:
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Assessment Scope
    row += 2
    ws[f"A{row}"] = "ASSESSMENT SCOPE"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:G{row}")

    row += 1
    scope_text = """This assessment evaluates the incident management framework and governance structure (A.5.24 focus).

INCLUDED:
✅ Incident management governance (policy, procedures, authority)
✅ CSIRT/SOC organizational structure and staffing
✅ Role definitions, responsibilities, RACI matrix
✅ Training and competency programs
✅ Tools and technology capabilities
✅ Integration with logging, monitoring, BC/DR

EXCLUDED:
❌ Incident detection and classification procedures (see S2)
❌ Response execution and effectiveness (see S3)
❌ Forensic evidence procedures (see S4)
❌ Post-incident learning processes (see S5)"""

    ws.merge_cells(f"A{row}:G{row+13}")
    ws[f"A{row}"] = scope_text
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 220

    # Color Legend
    row += 14
    ws[f"A{row}"] = "COLOR LEGEND"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:C{row}")

    row += 1
    legend = [
        ("Yellow", "User Input Required", "FFFFCC"),
        ("Light Green", "Calculated/Auto-filled", "C6EFCE"),
        ("White", "Read-Only Information", "FFFFFF"),
        ("Light Blue", "Section Header", "D8E4F8"),
        ("Light Red", "Gap Identified", "FFC7CE"),
    ]

    ws[f"A{row}"] = "Color"
    ws[f"B{row}"] = "Meaning"
    ws[f"C{row}"] = "Usage"
    for cell in [ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"]]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = styles["border"]
    
    row += 1
    for color_name, meaning, rgb in legend:
        ws[f"A{row}"] = color_name
        ws[f"A{row}"].fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = meaning
        ws[f"B{row}"].border = styles["border"]
        ws[f"C{row}"] = "Input cells" if color_name == "Yellow" else (
            "Formula cells" if color_name == "Light Green" else (
                "Instructions, labels" if color_name == "White" else (
                    "Section titles" if color_name == "Light Blue" else "Risk highlighting"
                )
            )
        )
        ws[f"C{row}"].border = styles["border"]
        row += 1

    # Assessment Workflow
    row += 2
    ws[f"A{row}"] = "ASSESSMENT WORKFLOW"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:G{row}")

    row += 1
    workflow_text = """1. Information Gathering (3-4 hours)
   • Review incident management policy and procedures
   • Collect organizational charts, job descriptions, RACI matrices
   • Gather training records and exercise logs
   • Inventory incident response tools and integrations

2. Assessment Completion (3-4 hours)
   • Complete Governance Assessment (25 questions)
   • Document Organizational Structure (30 questions)
   • Assess Training & Competency (25 questions)
   • Evaluate Tools & Technology (30 questions)
   • Complete Integration Assessment (25 questions)

3. Gap Analysis (1 hour)
   • Identify missing or incomplete elements
   • Prioritize gaps by risk level
   • Develop remediation recommendations

4. Evidence Collection (1-2 hours)
   • Screenshots and document exports
   • Training records and certificates
   • Tool configuration evidence

5. Quality Review (1-2 hours)
   • Self-assessment against policy requirements
   • Peer review with CSIRT team
   • Management review

6. Approval Workflow (1 week)
   • CSIRT Manager approval
   • CISO review and approval"""

    ws.merge_cells(f"A{row}:G{row+27}")
    ws[f"A{row}"] = workflow_text
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 380

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 3: GOVERNANCE ASSESSMENT SHEET
# ============================================================================

def create_governance_assessment(ws, styles):
    """Create the Governance Assessment sheet (25 questions)."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "Sheet 2: Governance Assessment"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A2"] = "Policy, Procedures, Authority, and Regulatory Compliance (25 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")

    # Column headers
    row = 4
    headers = ["Question_ID", "Section", "Question", "Answer", "Evidence_Reference", "Comments", "Gap_Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Questions (25 total across 6 sections)
    questions = [
        # Section A: Policy Documentation (Q1-Q5)
        ("Q1", "Policy Documentation", "Does [Organization] have a documented incident management policy?", "Yes/No/In Draft"),
        ("Q2", "Policy Documentation", "When was the current incident management policy approved? (DD.MM.YYYY)", "Date"),
        ("Q3", "Policy Documentation", "Who is the designated owner of the incident management policy? (Name, Title)", "Text"),
        ("Q4", "Policy Documentation", "How often is the incident management policy reviewed?", "Annually/Bi-Annually/Quarterly/Ad-Hoc/Not Defined"),
        ("Q5", "Policy Documentation", "When was the policy last reviewed? (DD.MM.YYYY)", "Date"),
        
        # Section B: Procedure Documentation (Q6-Q10)
        ("Q6", "Procedure Documentation", "Are incident response procedures documented?", "Yes - Comprehensive/Yes - Partial/No/In Development"),
        ("Q7", "Procedure Documentation", "In what format are incident response procedures documented? (Select all: Written/Flowcharts/Automated/Scenarios/Other)", "Text"),
        ("Q8", "Procedure Documentation", "Where are incident response procedures stored and how are they accessed?", "Text"),
        ("Q9", "Procedure Documentation", "Do procedures include up-to-date contact information (CSIRT, management, legal, external partners)?", "Yes - Current/Yes - Outdated/No"),
        ("Q10", "Procedure Documentation", "When were incident response procedures last updated? (DD.MM.YYYY)", "Date"),
        
        # Section C: Authority & Escalation (Q11-Q15)
        ("Q11", "Authority & Escalation", "Who has authority to declare a security incident? (Role/title)", "Text"),
        ("Q12", "Authority & Escalation", "Is there a documented escalation matrix defining when and to whom incidents are escalated?", "Yes - Comprehensive/Yes - Basic/No"),
        ("Q13", "Authority & Escalation", "What criteria trigger executive management notification?", "Text"),
        ("Q14", "Authority & Escalation", "What criteria trigger Board of Directors notification?", "Text"),
        ("Q15", "Authority & Escalation", "Are procedures defined for external notifications (regulators, law enforcement, customers, media)?", "Yes - Documented/Partially Documented/No"),
        
        # Section D: Regulatory Compliance (Q16-Q20)
        ("Q16", "Regulatory Compliance", "Have applicable regulatory incident notification requirements been identified?", "Yes - Documented/Partially/No"),
        ("Q17", "Regulatory Compliance", "(If GDPR applies) Is GDPR 72-hour breach notification procedure documented?", "Yes/No/N/A"),
        ("Q18", "Regulatory Compliance", "(If Swiss nDSG applies) Is Swiss nDSG breach notification procedure documented?", "Yes/No/N/A"),
        ("Q19", "Regulatory Compliance", "(If PCI DSS applies) Is PCI DSS incident reporting procedure documented?", "Yes/No/N/A"),
        ("Q20", "Regulatory Compliance", "Are sector-specific incident reporting requirements documented (FINMA, DORA, NIS2, HIPAA, etc.)?", "Yes/No/N/A"),
        
        # Section E: Exception Management (Q21-Q22)
        ("Q21", "Exception Management", "Is there a process for exceptions to incident management requirements?", "Yes/No"),
        ("Q22", "Exception Management", "Are current exceptions documented and approved?", "Yes/No/N/A (no exceptions)"),
        
        # Section F: Policy Governance (Q23-Q25)
        ("Q23", "Policy Governance", "How is the incident management policy communicated to staff? (Select all: Repository/Onboarding/Training/Other)", "Text"),
        ("Q24", "Policy Governance", "Can relevant staff easily access the incident management policy?", "Yes - Widely Accessible/Yes - Restricted Access/No"),
        ("Q25", "Policy Governance", "When is the next scheduled policy review? (DD.MM.YYYY)", "Date"),
    ]

    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]  # Answer cell (yellow)
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]  # Evidence (yellow)
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]  # Comments (yellow)
        
        # Gap formula in column G
        ws[f"G{row}"] = f'=IF(OR(D{row}="No", D{row}="Partial", D{row}="Informal"), "Yes", "No")'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 4: ORGANIZATIONAL STRUCTURE SHEET
# ============================================================================

def create_organizational_structure(ws, styles):
    """Create the Organizational Structure sheet (30 questions)."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "Sheet 3: Organizational Structure"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A2"] = "CSIRT/SOC Model, Roles, Staffing, and RACI (30 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")

    # Column headers
    row = 4
    headers = ["Question_ID", "Section", "Question", "Answer", "Evidence_Reference", "Comments", "Gap_Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Questions (30 total across 5 sections)
    questions = [
        # Section A: CSIRT/SOC Model (Q26-Q30)
        ("Q26", "CSIRT/SOC Model", "What incident response organizational model does [Organization] use?", "Dedicated CSIRT/Virtual CSIRT/Hybrid/Outsourced SOC/Coordinated"),
        ("Q27", "CSIRT/SOC Model", "When was the CSIRT/SOC team established? (MM.YYYY)", "Date"),
        ("Q28", "CSIRT/SOC Model", "What coverage does the CSIRT/SOC provide?", "24/7/365/Business Hours/Extended Hours/On-Call Rotation"),
        ("Q29", "CSIRT/SOC Model", "How many full-time equivalent (FTE) staff are dedicated to incident response?", "Number"),
        ("Q30", "CSIRT/SOC Model", "To whom does the CSIRT/SOC team report? (Title/Department)", "Text"),
        
        # Section B: Role Definitions (Q31-Q45 - 15 roles, showing first 5 here for brevity)
        ("Q31", "Role Definitions", "Incident Manager - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q32", "Role Definitions", "SOC Analyst Tier 1 - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q33", "Role Definitions", "SOC Analyst Tier 2 - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q34", "Role Definitions", "Forensic Specialist - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q35", "Role Definitions", "Malware Analyst - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q36", "Role Definitions", "Threat Intelligence Analyst - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q37", "Role Definitions", "Communication Coordinator - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q38", "Role Definitions", "Legal Representative - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q39", "Role Definitions", "HR Representative - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q40", "Role Definitions", "IT Operations Liaison - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q41", "Role Definitions", "Business Unit Liaison - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q42", "Role Definitions", "Third-Party Coordinator - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q43", "Role Definitions", "Executive Sponsor - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q44", "Role Definitions", "On-Call Engineer - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q45", "Role Definitions", "Other Roles - List any additional incident response roles:", "Text"),
        
        # Section C: RACI Matrix (Q46-Q48)
        ("Q46", "RACI Matrix", "Is there a documented RACI matrix for incident management roles?", "Yes - Comprehensive/Yes - Basic/No"),
        ("Q47", "RACI Matrix", "Overall, are roles and responsibilities clear and unambiguous?", "Yes - Very Clear/Mostly Clear/Some Ambiguity/Significant Confusion"),
        ("Q48", "RACI Matrix", "When was the RACI matrix last reviewed/updated? (DD.MM.YYYY)", "Date"),
        
        # Section D: Staffing Adequacy (Q49-Q52)
        ("Q49", "Staffing Adequacy", "Is current staffing adequate to support stated coverage?", "Yes/No/Barely Adequate"),
        ("Q50", "Staffing Adequacy", "What is the CSIRT/SOC annual turnover rate? (%)", "Percentage"),
        ("Q51", "Staffing Adequacy", "Is there succession planning for key incident response roles?", "Yes/Informal/No"),
        ("Q52", "Staffing Adequacy", "Have skill gaps in the incident response team been identified?", "Yes - Documented/Informally Identified/No"),
        
        # Section E: Management Visibility (Q53-Q55)
        ("Q53", "Management Visibility", "How frequently does CSIRT report to management (CISO, CIO, or Executive)?", "Weekly/Monthly/Quarterly/Ad-Hoc/Never"),
        ("Q54", "Management Visibility", "Are incident management performance metrics defined and tracked?", "Yes - Comprehensive/Yes - Basic/No"),
        ("Q55", "Management Visibility", "Is there a protocol for briefing the Board of Directors on material incidents?", "Yes/Informal/No/N/A"),
    ]

    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(OR(D{row}="No", D{row}="Vacant", D{row}="Barely Adequate"), "Yes", "No")'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 5: TRAINING & COMPETENCY SHEET
# ============================================================================

def create_training_competency(ws, styles):
    """Create the Training & Competency sheet (25 questions)."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "Sheet 4: Training & Competency"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A2"] = "Training Programs, Exercises, and Certifications (25 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")

    # Column headers
    row = 4
    headers = ["Question_ID", "Section", "Question", "Answer", "Evidence_Reference", "Comments", "Gap_Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Questions (25 total across 5 sections)
    questions = [
        # Section A: Training Program (Q56-Q60)
        ("Q56", "Training Program", "Does [Organization] have a documented incident response training program?", "Yes - Comprehensive/Yes - Basic/No"),
        ("Q57", "Training Program", "Do all staff receive annual security awareness training including incident reporting?", "Yes/Partial/No"),
        ("Q58", "Training Program", "What percentage of staff completed incident reporting awareness training in the last 12 months? (%)", "Percentage"),
        ("Q59", "Training Program", "Do CSIRT members receive specialized incident response training?", "Yes - Regular/Yes - Occasional/No"),
        ("Q60", "Training Program", "Is there a dedicated budget for incident response training (courses, certifications, conferences)?", "Yes - Adequate/Yes - Limited/No"),
        
        # Section B: Tabletop Exercises (Q61-Q65)
        ("Q61", "Tabletop Exercises", "How often does [Organization] conduct incident response tabletop exercises?", "Quarterly/Semi-Annually/Annually/Bi-Annually/Never"),
        ("Q62", "Tabletop Exercises", "When was the most recent tabletop exercise conducted? (DD.MM.YYYY)", "Date"),
        ("Q63", "Tabletop Exercises", "Do tabletop exercises cover a variety of incident types?", "Yes - Diverse Scenarios/Limited Variety/Single Scenario"),
        ("Q64", "Tabletop Exercises", "Who typically participates in tabletop exercises? (Select all: CSIRT/IT Ops/Legal/HR/Management/Executive/Board/External)", "Text"),
        ("Q65", "Tabletop Exercises", "Are lessons learned from tabletop exercises documented and acted upon?", "Yes - Systematically/Yes - Informally/No"),
        
        # Section C: Simulation & Testing (Q66-Q68)
        ("Q66", "Simulation & Testing", "Does [Organization] conduct full-scale incident response exercises (beyond tabletop)?", "Yes - Regularly/Yes - Occasionally/No"),
        ("Q67", "Simulation & Testing", "Does [Organization] use breach simulation or attack simulation tools?", "Yes/No"),
        ("Q68", "Simulation & Testing", "Are drills conducted specifically for regulatory breach notification (GDPR 72-hour, PCI DSS, etc.)?", "Yes/No"),
        
        # Section D: Competency & Certification (Q69-Q72)
        ("Q69", "Competency & Certification", "Are competency requirements defined for incident response roles?", "Yes/No"),
        ("Q70", "Competency & Certification", "Does [Organization] track security certifications for CSIRT members?", "Yes/Informal/No"),
        ("Q71", "Competency & Certification", "Are certifications current (not expired)?", "Yes - All Current/Some Expired/Not Tracked"),
        ("Q72", "Competency & Certification", "Is there a documented onboarding process for new CSIRT members?", "Yes - Formal Program/Informal/No"),
        
        # Section E: Training Effectiveness (Q73-Q75)
        ("Q73", "Training Effectiveness", "Is training effectiveness measured (beyond completion tracking)?", "Yes/No"),
        ("Q74", "Training Effectiveness", "Is feedback collected from training participants?", "Yes - Systematically/Occasionally/No"),
        ("Q75", "Training Effectiveness", "Have training gaps been identified through exercises or actual incidents?", "Yes - Documented/Informally Identified/No"),
        
        # Additional questions to reach 25
        ("Q76", "Training Records", "Are training records maintained and accessible for audit purposes?", "Yes/Partial/No"),
        ("Q77", "Training Records", "How long are training records retained? (Years)", "Number"),
        ("Q78", "External Training", "Have CSIRT members attended external training in the last 12 months?", "Yes/No"),
        ("Q79", "Certification Goals", "Are certification goals set for CSIRT members?", "Yes/No"),
        ("Q80", "Training Budget", "What percentage of security budget is allocated to incident response training? (%)", "Percentage"),
    ]

    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(OR(D{row}="No", D{row}<95, D{row}="Never"), "Yes", "No")'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 6: TOOLS & TECHNOLOGY SHEET
# ============================================================================

def create_tools_technology(ws, styles):
    """Create the Tools & Technology sheet (30 questions)."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "Sheet 5: Tools & Technology"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A2"] = "Ticketing, SIEM, Forensics, Communication, and Automation (30 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")

    # Column headers
    row = 4
    headers = ["Question_ID", "Section", "Question", "Answer", "Evidence_Reference", "Comments", "Gap_Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Questions (30 total across 7 sections)
    questions = [
        # Section A: Ticketing (Q76-Q79)
        ("Q81", "Ticketing", "Does [Organization] have a dedicated incident ticketing/tracking system?", "Yes - Dedicated IR/Yes - General IT/No"),
        ("Q82", "Ticketing", "What features does the incident ticketing system provide? (Select all: Creation/Severity/Timeline/Evidence/Tasks/Collaboration/Escalation/Integration/Reporting/Playbooks)", "Text"),
        ("Q83", "Ticketing", "How long are incident tickets retained? (Years)", "1 year/3 years/5 years/7+ years/Indefinitely/Not Defined"),
        ("Q84", "Ticketing", "Are access controls implemented for incident tickets (need-to-know basis)?", "Yes - Strict/Yes - Moderate/No - Open"),
        
        # Section B: SIEM (Q80-Q83)
        ("Q85", "SIEM", "Does [Organization] have a Security Information and Event Management (SIEM) system?", "Yes - Fully Operational/Yes - Partial/No"),
        ("Q86", "SIEM", "How many log sources feed into the SIEM? (Number)", "Number"),
        ("Q87", "SIEM", "Is the SIEM integrated with the incident ticketing system (automatic ticket creation)?", "Yes - Fully Automated/Yes - Semi-Automated/Manual/N/A"),
        ("Q88", "SIEM", "Does the SIEM/SOAR platform support automated playbooks for common incident types?", "Yes - Extensive/Yes - Limited/No/N/A"),
        
        # Section C: Forensics (Q84-Q87)
        ("Q89", "Forensics", "What digital forensic tools are available? (Select all: Disk imaging/Memory acquisition/Memory analysis/Network forensics/Log analysis/Malware analysis/Mobile/Cloud/Commercial suites/None)", "Text"),
        ("Q90", "Forensics", "Are CSIRT members trained in the use of forensic tools?", "Yes - All Members/Yes - Specialized Roles/Limited/No"),
        ("Q91", "Forensics", "Does [Organization] have a dedicated forensic workstation or lab?", "Yes - Dedicated Lab/Yes - Virtual/No"),
        ("Q92", "Forensics", "Is there secure storage for forensic evidence?", "Yes - Dedicated/Yes - General/No"),
        
        # Section D: Communication (Q88-Q90)
        ("Q93", "Communication", "What tools are used for internal incident response communication? (Select all: Email/Phone/IM/Video/War Room/IR Platform/Encrypted)", "Text"),
        ("Q94", "Communication", "Is there a secure channel for external communications (legal, law enforcement, regulators)?", "Yes/No"),
        ("Q95", "Communication", "How are on-call incident responders alerted?", "PagerDuty/VictorOps/Phone/SMS/Email/No System"),
        
        # Section E: Threat Intel (Q91-Q93)
        ("Q96", "Threat Intel", "Does [Organization] subscribe to threat intelligence feeds?", "Yes - Multiple/Yes - Single/No"),
        ("Q97", "Threat Intel", "Is threat intelligence integrated into SIEM for alerting?", "Yes - Automated/Yes - Manual/No/N/A"),
        ("Q98", "Threat Intel", "Is there a dedicated threat intelligence analyst role?", "Yes - Dedicated/Yes - Part-Time/No"),
        
        # Section F: Automation (Q94-Q96)
        ("Q99", "Automation", "What level of incident response automation exists?", "High (SOAR)/Moderate (Some Playbooks)/Low (Mostly Manual)/None"),
        ("Q100", "Automation", "How many documented/automated playbooks exist? (Number)", "Number"),
        ("Q101", "Automation", "Are playbooks regularly reviewed and updated?", "Yes - Annually/Yes - After Incidents/No/N/A"),
        
        # Section G: Tool Gaps (Q97-Q99)
        ("Q102", "Tool Gaps", "Are there critical tool gaps preventing effective incident response?", "Yes - Multiple/Yes - Some/No"),
        ("Q103", "Tool Gaps", "Are tool gaps being addressed (budget approved, procurement in progress)?", "Yes - Funded/Proposed/No"),
        ("Q104", "Tool Gaps", "Are there challenges integrating incident response tools?", "Yes - Significant/Yes - Moderate/No/N/A"),
        
        # Additional questions
        ("Q105", "Tool Inventory", "Is there a documented inventory of all incident response tools?", "Yes/No"),
        ("Q106", "Tool Licensing", "Are all incident response tools properly licensed?", "Yes/Some Unlicensed/Not Tracked"),
        ("Q107", "Tool Maintenance", "Are incident response tools regularly updated and maintained?", "Yes/Partial/No"),
        ("Q108", "Tool Training", "Do CSIRT members receive training on all critical tools?", "Yes/Partial/No"),
        ("Q109", "Tool Documentation", "Is tool documentation maintained and accessible?", "Yes/Partial/No"),
        ("Q110", "Tool Testing", "Are incident response tools tested regularly?", "Yes/Occasional/No"),
    ]

    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(OR(D{row}="No", D{row}="None", D{row}="N/A"), "Yes", "No")'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 7: INTEGRATION ASSESSMENT SHEET
# ============================================================================

def create_integration_assessment(ws, styles):
    """Create the Integration Assessment sheet (25 questions)."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "Sheet 6: Integration Assessment"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A2"] = "Integration with Logging, Monitoring, BC/DR, and Other Controls (25 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")

    # Column headers
    row = 4
    headers = ["Question_ID", "Section", "Question", "Answer", "Evidence_Reference", "Comments", "Gap_Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Questions (25 total across 6 sections)
    questions = [
        # Section A: Logging Integration (Q100-Q102)
        ("Q111", "Logging Integration", "Is incident management integrated with logging infrastructure (A.8.15)?", "Yes - Fully Integrated/Partially Integrated/No"),
        ("Q112", "Logging Integration", "Are logs readily available to CSIRT during incidents?", "Yes - Immediate Access/Yes - Delayed Access/No"),
        ("Q113", "Logging Integration", "Is log retention adequate for incident investigation?", "Yes/Insufficient for Some Systems/No"),
        
        # Section B: Monitoring Integration (Q103-Q105)
        ("Q114", "Monitoring Integration", "Is incident management integrated with security monitoring (A.8.16)?", "Yes - Fully Integrated/Partially Integrated/No"),
        ("Q115", "Monitoring Integration", "Is there a defined workflow from security alert to incident declaration?", "Yes - Documented/Informal/No"),
        ("Q116", "Monitoring Integration", "Is the false positive rate from monitoring tracked?", "Yes/No"),
        
        # Section C: User Reporting Integration (Q106-Q108)
        ("Q117", "User Reporting", "Is there a mechanism for users to report security events (A.6.8 integration)?", "Yes - Dedicated/Yes - General IT Helpdesk/No"),
        ("Q118", "User Reporting", "Are user-reported events integrated into incident ticketing system?", "Yes - Automatically/Yes - Manually/No"),
        ("Q119", "User Reporting", "Do users know how to report security incidents?", "Yes - Well Communicated/Somewhat/No"),
        
        # Section D: BC/DR Integration (Q109-Q111)
        ("Q120", "BC/DR Integration", "Is incident management integrated with Business Continuity / Disaster Recovery (A.5.29-30)?", "Yes - Documented/Informal/No"),
        ("Q121", "BC/DR Integration", "Are criteria defined for when an incident triggers BC/DR activation?", "Yes/No"),
        ("Q122", "BC/DR Integration", "Is CSIRT's role defined in BC/DR plans?", "Yes - Documented/Informal/No"),
        
        # Section E: Third-Party Coordination (Q112-Q114)
        ("Q123", "Third-Party Coordination", "Are procedures defined for incidents involving third parties (suppliers, MSSPs)?", "Yes - Documented/Informal/No"),
        ("Q124", "Third-Party Coordination", "(If MSSP/outsourced SOC) Are SLAs defined for incident detection and escalation?", "Yes/No/N/A"),
        ("Q125", "Third-Party Coordination", "Do suppliers know how to report security incidents to [Organization]?", "Yes - Defined in Contracts/Informal/No"),
        
        # Section F: Legal & Regulatory (Q115-Q117)
        ("Q126", "Legal & Regulatory", "Is Legal/Compliance involvement in incident management clearly defined?", "Yes - Documented/Informal/No"),
        ("Q127", "Legal & Regulatory", "(If GDPR applicable) Is the Data Protection Officer (DPO) integrated into incident management?", "Yes/No/N/A"),
        ("Q128", "Legal & Regulatory", "Are procedures defined for coordinating with law enforcement?", "Yes - Documented/Informal/No"),
        
        # Additional cross-integration questions
        ("Q129", "Asset Management", "Is incident management integrated with asset management (A.5.9)?", "Yes/Partial/No"),
        ("Q130", "Access Control", "Is incident management integrated with access control processes (A.5.15-18)?", "Yes/Partial/No"),
        ("Q131", "Change Management", "Is incident management integrated with change management (A.8.32)?", "Yes/Partial/No"),
        ("Q132", "Vulnerability Management", "Is incident management integrated with vulnerability management (A.8.8)?", "Yes/Partial/No"),
        ("Q133", "Patch Management", "Is incident management integrated with patch management (A.8.19)?", "Yes/Partial/No"),
        ("Q134", "Risk Assessment", "Are incident trends fed back into risk assessment processes (A.5.8)?", "Yes/Partial/No"),
        ("Q135", "Audit & Compliance", "Are incident records available for audit and compliance reviews?", "Yes/Partial/No"),
    ]

    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(OR(D{row}="No", D{row}="Informal"), "Yes", "No")'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 8: GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create the Gap Analysis sheet."""
    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "Sheet 7: Gap Analysis"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A2"] = "Gap Prioritization and Remediation Planning (60 Gap Capacity)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:J2")

    # Column headers
    row = 4
    headers = ["Gap_ID", "Assessment_Section", "Gap_Description", "Risk_Level", "Current_State", "Target_State", "Remediation_Action", "Owner", "Target_Date", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Example gaps (first 5 rows)
    example_gaps = [
        ("GAP-001", "Governance", "No RACI matrix documented", "High", "Roles informally understood", "RACI matrix created and approved", "Document RACI matrix, review with CSIRT, approve", "CSIRT Manager", "", "Open"),
        ("GAP-002", "Training", "Tabletop exercise overdue (>12 months)", "Medium", "Last exercise 18 months ago", "Quarterly exercises", "Schedule Q1 exercise, appoint facilitator", "CSIRT Manager", "", "Open"),
        ("GAP-003", "Tools", "No dedicated forensic workstation", "High", "No forensic lab", "Dedicated forensic workstation with write blockers", "Procure forensic workstation and tools", "IT Manager", "", "Open"),
        ("GAP-004", "Integration", "No BC/DR integration documented", "Medium", "Informal integration", "Formal integration in BC/DR plan", "Update BC/DR plan with CSIRT roles", "CISO", "", "Open"),
        ("GAP-005", "Staffing", "CSIRT understaffed for 24/7 coverage", "Critical", "2 FTE for 24/7", "5 FTE for proper shift coverage", "Hire 3 additional CSIRT analysts", "CISO", "", "Open"),
    ]

    row = 5
    for gap_data in example_gaps:
        for col_idx, value in enumerate(gap_data, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            if col_idx in [2, 3, 5, 6, 7, 8, 9]:  # Input cells
                cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            if col_idx == 4:  # Risk_Level
                if value == "Critical":
                    cell.fill = styles["gap_critical"]["fill"]
                elif value == "High":
                    cell.fill = styles["gap_high"]["fill"]
                elif value == "Medium":
                    cell.fill = styles["gap_medium"]["fill"]
                else:
                    cell.fill = styles["gap_low"]["fill"]
        row += 1

    # Add empty rows for additional gaps (total 60 capacity)
    for i in range(55):
        ws[f"A{row}"] = f"GAP-{str(row-4).zfill(3)}"
        for col_idx in range(2, 11):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Summary section
    row += 2
    ws[f"A{row}"] = "GAP SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:D{row}")

    row += 1
    summary_labels = ["Total Gaps:", "Critical:", "High:", "Medium:", "Low:", "", "Open:", "In Progress:", "Resolved:", "Accepted:"]
    for label in summary_labels:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        if label:  # Don't create formula for empty row
            if label == "Total Gaps:":
                ws[f"B{row}"] = "=COUNTA(A5:A64)"
            elif label in ["Critical:", "High:", "Medium:", "Low:"]:
                ws[f"B{row}"] = f'=COUNTIF(D5:D64, "{label[:-1]}")'
            elif label in ["Open:", "In Progress:", "Resolved:", "Accepted:"]:
                ws[f"B{row}"] = f'=COUNTIF(J5:J64, "{label[:-1]}")'
            ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 9: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """Create the Evidence Register sheet."""
    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "Sheet 8: Evidence Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A2"] = "Audit Evidence Tracking (100 Items Capacity)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:H2")

    # Column headers
    row = 4
    headers = ["Evidence_ID", "Evidence_Type", "Description", "Related_Section", "Storage_Location", "Date_Collected", "Collected_By", "Verification_Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Example evidence (first 5 rows)
    example_evidence = [
        ("EV-001", "Policy Document", "ISMS-POL-A.5.24-28", "Governance", "SharePoint:/Policies/", datetime.now().strftime("%d.%m.%Y"), "CSIRT Manager", "Verified"),
        ("EV-002", "Org Chart", "CSIRT Organizational Chart", "Structure", "/Evidence/S1/org_chart.pdf", datetime.now().strftime("%d.%m.%Y"), "HR", "Verified"),
        ("EV-003", "Training Records", "2025 Security Awareness Training Completion", "Training", "HR System Report", datetime.now().strftime("%d.%m.%Y"), "HR", "Verified"),
        ("EV-004", "Tool Screenshot", "Incident Ticketing System Dashboard", "Tools", "/Evidence/S1/ticketing_system.png", datetime.now().strftime("%d.%m.%Y"), "SOC Manager", "Pending"),
        ("EV-005", "RACI Matrix", "Incident Response RACI Matrix v1.0", "Structure", "SharePoint:/CSIRT/RACI.xlsx", datetime.now().strftime("%d.%m.%Y"), "CSIRT Manager", "Verified"),
    ]

    row = 5
    for ev_data in example_evidence:
        for col_idx, value in enumerate(ev_data, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            if col_idx in [2, 3, 4, 5, 6, 7, 8]:  # Input cells
                cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Add empty rows (total 100 capacity)
    for i in range(95):
        ws[f"A{row}"] = f"EV-{str(row-4).zfill(3)}"
        for col_idx in range(2, 9):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 10: DASHBOARD SHEET
# ============================================================================

def create_dashboard(ws, styles):
    """Create the Dashboard sheet with maturity scoring."""
    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "Sheet 9: Dashboard"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A2"] = "Executive Summary - Framework Maturity Assessment"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:H2")

    # Overall Maturity Score
    row = 4
    ws[f"A{row}"] = "OVERALL FRAMEWORK MATURITY"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:D{row}")

    row += 1
    ws[f"A{row}"] = "Overall Maturity Score (%)"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "=100 - ((COUNTIF('Governance Assessment'!G:G,\"Yes\") + COUNTIF('Organizational Structure'!G:G,\"Yes\") + COUNTIF('Training & Competency'!G:G,\"Yes\") + COUNTIF('Tools & Technology'!G:G,\"Yes\") + COUNTIF('Integration Assessment'!G:G,\"Yes\")) / 135 * 100)"
    ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]
    ws[f"B{row}"].number_format = "0.0"

    row += 1
    ws[f"A{row}"] = "Maturity Level"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = '=IF(B5<40,"Level 1 (Initial)",IF(B5<60,"Level 2 (Developing)",IF(B5<75,"Level 3 (Defined)",IF(B5<90,"Level 4 (Managed)","Level 5 (Optimizing)"))))'
    ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]

    # Domain Scores
    row += 3
    ws[f"A{row}"] = "DOMAIN SCORES"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:D{row}")

    row += 1
    ws[f"A{row}"] = "Domain"
    ws[f"B{row}"] = "Score (%)"
    ws[f"C{row}"] = "Target"
    ws[f"D{row}"] = "Status"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f"{col}{row}"].border = styles["border"]

    domains = [
        ("Governance", "=100 - (COUNTIF('Governance Assessment'!G:G,\"Yes\") / 25 * 100)", "90%"),
        ("Organizational Structure", "=100 - (COUNTIF('Organizational Structure'!G:G,\"Yes\") / 30 * 100)", "85%"),
        ("Training & Competency", "=100 - (COUNTIF('Training & Competency'!G:G,\"Yes\") / 25 * 100)", "90%"),
        ("Tools & Technology", "=100 - (COUNTIF('Tools & Technology'!G:G,\"Yes\") / 30 * 100)", "80%"),
        ("Integration", "=100 - (COUNTIF('Integration Assessment'!G:G,\"Yes\") / 25 * 100)", "85%"),
    ]

    row += 1
    for domain_name, formula, target in domains:
        ws[f"A{row}"] = domain_name
        ws[f"B{row}"] = formula
        ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]
        ws[f"B{row}"].number_format = "0.0"
        ws[f"C{row}"] = target
        ws[f"D{row}"] = "=IF(B" + str(row) + ">=VALUE(LEFT(C" + str(row) + ",LEN(C" + str(row) + ")-1)),\"On Track\",\"Needs Improvement\")"
        ws[f"D{row}"].fill = styles["calculated_cell"]["fill"]
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = styles["border"]
        row += 1

    # Key Metrics
    row += 2
    ws[f"A{row}"] = "KEY METRICS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:D{row}")

    row += 1
    metrics = [
        ("CSIRT Staffing (FTE)", "='Organizational Structure'!D29"),
        ("Coverage Model", "='Organizational Structure'!D28"),
        ("Training Completion Rate", "='Training & Competency'!D58"),
        ("Tabletop Exercise Frequency", "='Training & Competency'!D61"),
        ("Critical Gaps", "='Gap Analysis'!B72"),
        ("High Gaps", "='Gap Analysis'!B73"),
    ]

    for metric_label, metric_formula in metrics:
        ws[f"A{row}"] = metric_label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = metric_formula
        ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20

    ws.freeze_panes = "A4"


# ============================================================================
# SECTION 11: APPROVAL SIGN-OFF SHEET
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create the Approval Sign-Off sheet."""
    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "Sheet 10: Approval & Sign-Off"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Assessment Summary
    row = 3
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")

    summary_fields = [
        ("Assessment Document", "ISMS-IMP-A.5.24-28.S1 - Framework Assessment"),
        ("Assessment Period", ""),  # User input
        ("Overall Maturity Level", "=Dashboard!B6"),
        ("Overall Maturity Score", "=Dashboard!B5"),
        ("Assessment Status", ""),  # Dropdown
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if label in ["Assessment Period", "Assessment Status"]:
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        else:
            ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Assessment Completed By
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    completion_fields = ["Name", "Role/Title", "Department", "Email", "Date"]
    row += 1
    for field in completion_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Reviewed By (CISO)
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    review_fields = ["Name", "Date", "Approval Decision", "Signature"]
    row += 1
    for field in review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]
        row += 1

    # Next Review Date
    row += 2
    ws[f"A{row}"] = "Next Review Date:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]
    ws[f"B{row}"].alignment = styles["input_cell"]["alignment"]

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.5.24-28.S1 - Framework Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.5.24: Planning & Preparation")
    logger.info("=" * 80)

    wb = create_workbook()
    styles = setup_styles()

    logger.info("\n[1/10] Creating Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)

    logger.info("[2/10] Creating Governance Assessment sheet (25 questions)...")
    create_governance_assessment(wb["Governance Assessment"], styles)

    logger.info("[3/10] Creating Organizational Structure sheet (30 questions)...")
    create_organizational_structure(wb["Organizational Structure"], styles)

    logger.info("[4/10] Creating Training & Competency sheet (25 questions)...")
    create_training_competency(wb["Training & Competency"], styles)

    logger.info("[5/10] Creating Tools & Technology sheet (30 questions)...")
    create_tools_technology(wb["Tools & Technology"], styles)

    logger.info("[6/10] Creating Integration Assessment sheet (25 questions)...")
    create_integration_assessment(wb["Integration Assessment"], styles)

    logger.info("[7/10] Creating Gap Analysis sheet...")
    create_gap_analysis(wb["Gap Analysis"], styles)

    logger.info("[8/10] Creating Evidence Register sheet...")
    create_evidence_register(wb["Evidence Register"], styles)

    logger.info("[9/10] Creating Dashboard sheet...")
    create_dashboard(wb["Dashboard"], styles)

    logger.info("[10/10] Creating Approval Sign-Off sheet...")
    create_approval_signoff(wb["Approval Sign-Off"], styles)

    filename = f"ISMS-IMP-A.5.24-28.S1_Framework_Assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    logger.info(f"\n✅ SUCCESS: {filename}")
    logger.info("\nWorkbook Structure:")
    logger.info("  • 10 sheets (Instructions through Approval)")
    logger.info("  • 135 assessment questions (25+30+25+30+25)")
    logger.info("  • 60 gap analysis capacity")
    logger.info("  • 100 evidence register capacity")
    logger.info("  • Automated maturity scoring (Levels 1-5)")
    logger.info("\nNext steps:")
    logger.info("  1) Complete document information in Instructions & Legend")
    logger.info("  2) Conduct stakeholder interviews (CSIRT, Legal, HR, IT Ops)")
    logger.info("  3) Fill yellow cells in assessment sheets (Governance through Integration)")
    logger.info("  4) Review automated gap identification in Gap Analysis")
    logger.info("  5) Collect and link evidence in Evidence Register")
    logger.info("  6) Review Dashboard for maturity scoring")
    logger.info("  7) Complete Approval Sign-Off workflow")
    logger.info("\nEstimated completion time: 8-12 hours")
    logger.info("\n" + "=" * 80)


if __name__ == "__main__":
    main()


# ============================================================================
# END OF SCRIPT
# ============================================================================

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
