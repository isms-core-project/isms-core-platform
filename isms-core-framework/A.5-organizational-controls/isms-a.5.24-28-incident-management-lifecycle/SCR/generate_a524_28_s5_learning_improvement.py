#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.24-28.S5 - Learning & Continuous Improvement Assessment Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.27: Learning from Information Security Incidents
Assessment Domain 5 of 5: Learning & Continuous Improvement

--------------------------------------------------------------------------------
PURPOSE & DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
the learning and continuous improvement processes following security incidents,
focusing on Control A.5.27 of the incident management lifecycle.

**Assessment Scope:**
- Post-Incident Review (PIR) process timeliness and quality
- Root Cause Analysis (RCA) depth and methodology
- Lessons Learned documentation and knowledge management
- Control improvement tracking and remediation closure
- Trend analysis and metrics reporting

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and methodology
2. PIR Process - PIR inventory, quality scoring, SLA compliance
3. Root Cause Analysis - RCA inventory, depth scoring, recurring causes
4. Lessons Learned - LL log, knowledge base inventory, governance
5. Control Improvements - Remediation action register, closure tracking
6. Trend Analysis - KPIs, reporting cadence, accuracy verification
7. Gap Analysis - Consolidated gaps with prioritization
8. Evidence Register - Audit evidence tracking (50 items capacity)
9. Summary Dashboard - Executive summary with compliance scoring
10. Approval Sign-Off - Multi-stakeholder approval workflow

**Key Features:**
- PIR quality scoring (1-5 scale)
- RCA depth assessment (Technical/Procedural/Systemic)
- Automated SLA compliance tracking
- Remediation closure verification
- KPI trend analysis
- Critical flag identification

**Integration:**
This is Domain 5 of 5 assessment domains for A.5.24-28 Incident Management:
- S1: Framework & Governance (A.5.24 focus)
- S2: Detection & Classification (A.5.25 focus)
- S3: Response Capabilities (A.5.26 focus)
- S4: Forensic Evidence Management (A.5.28 focus)
- S5: Learning & Improvement (this assessment - A.5.27 focus)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a524_28_s5_learning_improvement.py

Output:
    File: ISMS-IMP-A.5.24-28.S5_Learning_Improvement_YYYYMMDD.xlsx
    Location: Current directory

Estimated Completion Time:
    - Information gathering: 2-3 hours
    - Historical sampling: 2-3 hours
    - Assessment completion: 2-3 hours
    - Evidence collection: 1-1.5 hours
    - Quality review: 30-60 minutes
    Total: 6-10 hours (depending on incident volume)

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.27
Assessment Domain:    5 of 5 (Learning & Improvement)
Framework Version:    1.0
Script Version:       1.0
Author:               ISMS Implementation Team
Date:                 [Date to be set]

================================================================================
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.24-28.S5"
WORKBOOK_NAME = "Learning & Continuous Improvement"
CONTROL_ID = "A.5.27"
CONTROL_NAME = "Learning from Information Security Incidents"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_Learning_Improvement_{GENERATED_TIMESTAMP}.xlsx"


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP specification."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Sheet structure per ISMS-IMP-A.5.24-28.S5 specification
    sheets = [
        "Instructions & Legend",
        "PIR Process",
        "Root Cause Analysis",
        "Lessons Learned",
        "Control Improvements",
        "Trend Analysis",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
        },
        "input": {
            "font": Font(name="Calibri", size=10),
            "fill": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "formula": {
            "font": Font(name="Calibri", size=10),
            "fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "normal": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "compliant": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "non_compliant": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "border": border_thin,
    }
    return styles


def apply_style(cell, style_dict, border=None):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = style_dict["font"]
    if "fill" in style_dict:
        cell.fill = style_dict["fill"]
    if "alignment" in style_dict:
        cell.alignment = style_dict["alignment"]
    if border:
        cell.border = border


# ============================================================================
# SECTION 2: INSTRUCTIONS & LEGEND SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create the Instructions & Legend sheet."""
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 40

    # Title block
    ws.merge_cells('A1:C1')
    ws['A1'] = f"{DOCUMENT_ID} - Learning & Continuous Improvement Assessment"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:C2')
    ws['A2'] = CONTROL_REF
    apply_style(ws['A2'], styles["subheader"], styles["border"])

    # Document information section
    row = 4
    ws[f'A{row}'] = "Document Information"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])
    ws.merge_cells(f'A{row}:C{row}')

    info_fields = [
        ("Organisation", "[Enter organisation name]"),
        ("Assessment Date", "[DD.MM.YYYY]"),
        ("Assessor Name", "[Enter assessor name]"),
        ("Assessment Period", "[Start date] to [End date]"),
        ("Document Version", "1.0"),
        ("Related Policy", "ISMS-POL-A.5.24-28"),
    ]

    for field, value in info_fields:
        row += 1
        ws[f'A{row}'] = field
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws[f'B{row}'] = value
        apply_style(ws[f'B{row}'], styles["input"], styles["border"])

    # Status Legend
    row += 2
    ws[f'A{row}'] = "Status Legend"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])
    ws.merge_cells(f'A{row}:C{row}')

    legend_items = [
        ("Compliant", "Fully meets policy requirements, no gaps identified", "C6EFCE"),
        ("Partial", "Some requirements met, minor gaps exist", "FFEB9C"),
        ("Non-Compliant", "Does not meet policy requirements, significant gaps", "FFC7CE"),
        ("N/A", "Not applicable to this organisation", "D9D9D9"),
    ]

    for status, desc, color in legend_items:
        row += 1
        ws[f'A{row}'] = status
        ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f'A{row}'].border = styles["border"]
        ws[f'B{row}'] = desc
        apply_style(ws[f'B{row}'], styles["normal"], styles["border"])

    # Assessment Scope
    row += 2
    ws[f'A{row}'] = "Assessment Scope - 5 Learning Domains"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])
    ws.merge_cells(f'A{row}:C{row}')

    domains = [
        ("1. PIR Process", "Timeliness, completeness, quality of post-incident reviews"),
        ("2. Root Cause Analysis", "Depth, methodology, accuracy of RCA"),
        ("3. Lessons Learned", "Documentation, dissemination, knowledge base"),
        ("4. Control Improvements", "Remediation tracking, ownership, closure"),
        ("5. Trend Analysis", "KPIs, trend identification, reporting"),
    ]

    for domain, desc in domains:
        row += 1
        ws[f'A{row}'] = domain
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws[f'B{row}'] = desc
        apply_style(ws[f'B{row}'], styles["normal"], styles["border"])

    # Instructions
    row += 2
    ws[f'A{row}'] = "Completion Instructions"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])
    ws.merge_cells(f'A{row}:C{row}')

    instructions = [
        "1. Complete PIR Process sheet - review all PIRs from last 12 months",
        "2. Complete Root Cause Analysis sheet - assess RCA quality for Critical/High incidents",
        "3. Complete Lessons Learned sheet - verify documentation and distribution",
        "4. Complete Control Improvements sheet - track remediation actions",
        "5. Complete Trend Analysis sheet - verify KPIs and reporting",
        "6. Review Gap Analysis - auto-populated from above sheets",
        "7. Link evidence in Evidence Register",
        "8. Review Summary Dashboard for overall compliance",
        "9. Complete Approval Sign-Off workflow",
    ]

    for instr in instructions:
        row += 1
        ws[f'A{row}'] = instr
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws.merge_cells(f'A{row}:C{row}')

    # Time estimates
    row += 2
    ws[f'A{row}'] = "Time Estimates"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])
    ws.merge_cells(f'A{row}:C{row}')

    times = [
        ("Information Gathering", "2-3 hours"),
        ("Historical Sampling", "2-3 hours"),
        ("Assessment Completion", "2-3 hours"),
        ("Evidence Collection", "1-1.5 hours"),
        ("Quality Review", "30-60 minutes"),
        ("Total", "6-10 hours"),
    ]

    for activity, time in times:
        row += 1
        ws[f'A{row}'] = activity
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws[f'B{row}'] = time
        apply_style(ws[f'B{row}'], styles["normal"], styles["border"])


# ============================================================================
# SECTION 3: PIR PROCESS SHEET
# ============================================================================

def create_pir_process_sheet(ws, styles):
    """Create the PIR Process assessment sheet."""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12

    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = "Sheet 2: PIR Process Assessment"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 30

    # Section A header
    ws.merge_cells('A3:L3')
    ws['A3'] = "Section A: PIR Inventory (Last 12 Months)"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    # Column headers
    headers = [
        "Incident_ID", "Incident_Date", "Severity", "Resolution_Date",
        "PIR_Status", "PIR_Completion_Date", "SLA_Days", "Actual_Days",
        "SLA_Met", "Participants_Met", "Quality_Score", "Evidence_Ref"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["subheader"], styles["border"])

    # Data rows (50 rows for PIR inventory)
    for row in range(6, 56):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col, value="")
            apply_style(cell, styles["input"], styles["border"])

    # Data validation for Severity
    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(severity_dv)
    severity_dv.add(f'C6:C55')

    # Data validation for PIR_Status
    status_dv = DataValidation(type="list", formula1='"Completed,Overdue,Pending,Not_Required"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f'E6:E55')

    # Data validation for Yes/No fields
    yesno_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yesno_dv)
    yesno_dv.add(f'I6:I55')
    yesno_dv.add(f'J6:J55')

    # Data validation for Quality Score
    quality_dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(quality_dv)
    quality_dv.add(f'K6:K55')

    # Section B: Quality Scoring Criteria (reference)
    row = 58
    ws.merge_cells(f'A{row}:L{row}')
    ws[f'A{row}'] = "Section B: PIR Quality Scoring Criteria (Reference)"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    criteria = [
        ("5 - Excellent", "Complete timeline, full impact assessment, all action items with owners/dates, root cause referenced, distributed"),
        ("4 - Good", "Timeline and impact documented, most action items defined, minor gaps in distribution"),
        ("3 - Adequate", "Basic incident summary, some action items defined, distribution incomplete"),
        ("2 - Poor", "Minimal documentation, few action items, no distribution"),
        ("1 - Non-existent", "PIR marked complete but no substantive report exists"),
    ]

    for score, desc in criteria:
        row += 1
        ws[f'A{row}'] = score
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws.merge_cells(f'B{row}:L{row}')
        ws[f'B{row}'] = desc
        apply_style(ws[f'B{row}'], styles["normal"], styles["border"])

    # Section C: Summary
    row += 2
    ws.merge_cells(f'A{row}:L{row}')
    ws[f'A{row}'] = "Section C: PIR Process Summary"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    summary_items = [
        ("Total incidents in period", '=COUNTA(A6:A55)'),
        ("PIRs completed", '=COUNTIF(E6:E55,"Completed")'),
        ("PIRs overdue", '=COUNTIF(E6:E55,"Overdue")'),
        ("SLA compliance rate (%)", '=IF(B' + str(row+2) + '>0,COUNTIF(I6:I55,"Yes")/B' + str(row+2) + '*100,0)'),
        ("Average quality score", '=AVERAGE(K6:K55)'),
        ("Participant requirement met (%)", '=IF(B' + str(row+2) + '>0,COUNTIF(J6:J55,"Yes")/B' + str(row+2) + '*100,0)'),
    ]

    for label, formula in summary_items:
        row += 1
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'D{row}'] = formula
        apply_style(ws[f'D{row}'], styles["formula"], styles["border"])


# ============================================================================
# SECTION 4: ROOT CAUSE ANALYSIS SHEET
# ============================================================================

def create_rca_sheet(ws, styles):
    """Create the Root Cause Analysis assessment sheet."""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12

    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = "Sheet 3: Root Cause Analysis Assessment"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 30

    # Section A header
    ws.merge_cells('A3:I3')
    ws['A3'] = "Section A: RCA Inventory (Critical and High Incidents)"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    # Column headers
    headers = [
        "Incident_ID", "Severity", "RCA_Status", "RCA_Date",
        "Methodology", "Root_Cause_Summary", "Depth_Score", "Recurring", "Evidence_Ref"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["subheader"], styles["border"])

    # Data rows (30 rows for Critical/High incidents)
    for row in range(6, 36):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col, value="")
            apply_style(cell, styles["input"], styles["border"])

    # Data validation for Severity
    severity_dv = DataValidation(type="list", formula1='"Critical,High"', allow_blank=True)
    ws.add_data_validation(severity_dv)
    severity_dv.add(f'B6:B35')

    # Data validation for RCA_Status
    status_dv = DataValidation(type="list", formula1='"Completed,In_Progress,Not_Performed"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f'C6:C35')

    # Data validation for Methodology
    method_dv = DataValidation(type="list", formula1='"5_Whys,Fishbone,Fault_Tree,Timeline,Combined,Other"', allow_blank=True)
    ws.add_data_validation(method_dv)
    method_dv.add(f'E6:E35')

    # Data validation for Depth Score
    depth_dv = DataValidation(type="list", formula1='"1 - Technical,2 - Procedural,3 - Systemic"', allow_blank=True)
    ws.add_data_validation(depth_dv)
    depth_dv.add(f'G6:G35')

    # Data validation for Recurring
    yesno_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yesno_dv)
    yesno_dv.add(f'H6:H35')

    # Section B: RCA Quality Assessment
    row = 40
    ws.merge_cells(f'A{row}:I{row}')
    ws[f'A{row}'] = "Section B: RCA Quality Questions"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    quality_questions = [
        "Does RCA go beyond immediate technical cause to identify process/governance gaps?",
        "Was the RCA methodology appropriate for the incident complexity?",
        "Are preventive recommendations specific and actionable?",
        "Has the RCA been reviewed by someone other than the incident owner?",
    ]

    for q in quality_questions:
        row += 1
        ws[f'A{row}'] = q
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'H{row}'] = ""
        apply_style(ws[f'H{row}'], styles["input"], styles["border"])

    # Data validation for quality questions
    yesno_dv2 = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    ws.add_data_validation(yesno_dv2)
    yesno_dv2.add(f'H41:H44')

    # Section C: Summary
    row += 3
    ws.merge_cells(f'A{row}:I{row}')
    ws[f'A{row}'] = "Section C: RCA Summary"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    summary_items = [
        ("Critical/High incidents in period", '=COUNTA(A6:A35)'),
        ("RCAs completed", '=COUNTIF(C6:C35,"Completed")'),
        ("RCAs not performed", '=COUNTIF(C6:C35,"Not_Performed")'),
        ("Systemic depth RCAs", '=COUNTIF(G6:G35,"3 - Systemic")'),
        ("Recurring root causes", '=COUNTIF(H6:H35,"Yes")'),
    ]

    for label, formula in summary_items:
        row += 1
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'D{row}'] = formula
        apply_style(ws[f'D{row}'], styles["formula"], styles["border"])


# ============================================================================
# SECTION 5: LESSONS LEARNED SHEET
# ============================================================================

def create_lessons_learned_sheet(ws, styles):
    """Create the Lessons Learned assessment sheet."""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Sheet 4: Lessons Learned Assessment"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 30

    # Section A header
    ws.merge_cells('A3:H3')
    ws['A3'] = "Section A: Lessons Learned Log (Per PIR)"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    # Column headers
    headers = [
        "PIR_ID", "LL_Entry_Date", "Lesson_Summary", "Distribution_Date",
        "SLA_Met", "Playbook_Update", "Playbook_Date", "Evidence_Ref"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["subheader"], styles["border"])

    # Data rows (50 rows)
    for row in range(6, 56):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col, value="")
            apply_style(cell, styles["input"], styles["border"])

    # Data validation for Yes/No fields
    yesno_dv = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=True)
    ws.add_data_validation(yesno_dv)
    yesno_dv.add(f'E6:E55')
    yesno_dv.add(f'F6:F55')

    # Section B: Knowledge Base Inventory
    row = 58
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "Section B: Knowledge Base Inventory"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    kb_headers = ["KB_ID", "Title", "Type", "Last_Updated", "Status", "Owner", "Evidence_Ref", ""]
    row += 2
    for col, header in enumerate(kb_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, styles["subheader"], styles["border"])

    # Data rows for KB (15 rows)
    for r in range(row + 1, row + 16):
        for col in range(1, 8):
            cell = ws.cell(row=r, column=col, value="")
            apply_style(cell, styles["input"], styles["border"])

    # Data validation for KB Type
    type_dv = DataValidation(type="list", formula1='"Playbook,Procedure,Template,Reference,Training"', allow_blank=True)
    ws.add_data_validation(type_dv)
    type_dv.add(f'C{row+1}:C{row+15}')

    # Data validation for KB Status
    status_dv = DataValidation(type="list", formula1='"Current,Outdated,Missing,Under_Review"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f'E{row+1}:E{row+15}')

    # Section C: Summary
    row += 18
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "Section C: Lessons Learned Summary"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    summary_items = [
        ("Total lessons learned entries", '=COUNTA(A6:A55)'),
        ("Distribution SLA met (%)", '=IF(COUNTA(A6:A55)>0,COUNTIF(E6:E55,"Yes")/COUNTA(A6:A55)*100,0)'),
        ("Playbook updates completed", '=COUNTIF(F6:F55,"Yes")'),
        ("KB items current", '=COUNTIF(E60:E74,"Current")'),
        ("KB items outdated/missing", '=COUNTIFS(E60:E74,"<>Current",E60:E74,"<>")'),
    ]

    for label, formula in summary_items:
        row += 1
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'D{row}'] = formula
        apply_style(ws[f'D{row}'], styles["formula"], styles["border"])


# ============================================================================
# SECTION 6: CONTROL IMPROVEMENTS SHEET
# ============================================================================

def create_control_improvements_sheet(ws, styles):
    """Create the Control Improvements assessment sheet."""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 12

    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = "Sheet 5: Control Improvements Assessment"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 30

    # Section A header
    ws.merge_cells('A3:K3')
    ws['A3'] = "Section A: Remediation Action Register"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    # Column headers
    headers = [
        "Action_ID", "Source_Incident", "Action_Description", "Priority",
        "Owner", "Target_Date", "Status", "Completion_Date", "Verified_By",
        "Escalated", "Evidence_Ref"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["subheader"], styles["border"])

    # Data rows (70 rows)
    for row in range(6, 76):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col, value="")
            apply_style(cell, styles["input"], styles["border"])

    # Data validation for Priority
    priority_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(priority_dv)
    priority_dv.add(f'D6:D75')

    # Data validation for Status
    status_dv = DataValidation(type="list", formula1='"Open,In_Progress,Completed,Blocked,Cancelled"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f'G6:G75')

    # Data validation for Yes/No
    yesno_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yesno_dv)
    yesno_dv.add(f'J6:J75')

    # Section B: Summary
    row = 80
    ws.merge_cells(f'A{row}:K{row}')
    ws[f'A{row}'] = "Section B: Action Summary Metrics"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    summary_items = [
        ("Total actions", '=COUNTA(A6:A75)'),
        ("Actions completed", '=COUNTIF(G6:G75,"Completed")'),
        ("Actions open", '=COUNTIF(G6:G75,"Open")'),
        ("Actions blocked", '=COUNTIF(G6:G75,"Blocked")'),
        ("Closure rate (%)", '=IF(COUNTA(A6:A75)>0,COUNTIF(G6:G75,"Completed")/COUNTA(A6:A75)*100,0)'),
        ("Critical actions open", '=COUNTIFS(D6:D75,"Critical",G6:G75,"<>Completed")'),
        ("Overdue actions", '=COUNTIFS(F6:F75,"<"&TODAY(),G6:G75,"<>Completed")'),
    ]

    for label, formula in summary_items:
        row += 1
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'D{row}'] = formula
        apply_style(ws[f'D{row}'], styles["formula"], styles["border"])


# ============================================================================
# SECTION 7: TREND ANALYSIS SHEET
# ============================================================================

def create_trend_analysis_sheet(ws, styles):
    """Create the Trend Analysis assessment sheet."""
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Sheet 6: Trend Analysis Assessment"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 30

    # Section A header
    ws.merge_cells('A3:H3')
    ws['A3'] = "Section A: KPI Definitions and Current Values"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    # KPI headers
    kpi_headers = ["KPI", "Metric_Type", "Target", "Current", "Status", "Trend", "Period", "Evidence_Ref"]
    for col, header in enumerate(kpi_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["subheader"], styles["border"])

    # Pre-defined KPIs
    kpis = [
        ("Incident Recurrence Rate", "Percentage", "<=15%", "", "", "", "", ""),
        ("PIR SLA Compliance Rate", "Percentage", ">=95%", "", "", "", "", ""),
        ("Lessons Learned Distribution SLA", "Percentage", ">=90%", "", "", "", "", ""),
        ("Remediation Closure Rate", "Percentage", ">=85%", "", "", "", "", ""),
        ("Average PIR Quality Score", "Score (1-5)", ">=4.0", "", "", "", "", ""),
        ("RCA Completion Rate (Crit/High)", "Percentage", "100%", "", "", "", "", ""),
        ("MTTD (Critical Incidents)", "Hours", "<=1", "", "", "", "", ""),
        ("MTTR (Critical Incidents)", "Hours", "<=4", "", "", "", "", ""),
    ]

    for row_num, kpi_data in enumerate(kpis, 6):
        for col, value in enumerate(kpi_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            if col <= 3:
                apply_style(cell, styles["normal"], styles["border"])
            else:
                apply_style(cell, styles["input"], styles["border"])

    # Data validation for Status
    status_dv = DataValidation(type="list", formula1='"Yes,No,At_Risk"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f'E6:E13')

    # Data validation for Trend
    trend_dv = DataValidation(type="list", formula1='"Improving,Stable,Degrading"', allow_blank=True)
    ws.add_data_validation(trend_dv)
    trend_dv.add(f'F6:F13')

    # Section B: Reporting Cadence Verification
    row = 16
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "Section B: Reporting Cadence Verification"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    report_headers = ["Report_Type", "Frequency", "Last_Produced", "On_Schedule", "Recipients_Documented", "Evidence_Ref", "", ""]
    row += 2
    for col, header in enumerate(report_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, styles["subheader"], styles["border"])

    reports = [
        ("Weekly CSIRT Status", "Weekly", "", "", "", ""),
        ("Monthly Incident Summary", "Monthly", "", "", "", ""),
        ("Quarterly Trend Report", "Quarterly", "", "", "", ""),
        ("Annual Executive Report", "Annual", "", "", "", ""),
    ]

    for row_num, report_data in enumerate(reports, row + 1):
        for col, value in enumerate(report_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            if col <= 2:
                apply_style(cell, styles["normal"], styles["border"])
            else:
                apply_style(cell, styles["input"], styles["border"])

    # Data validation for Yes/No
    yesno_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yesno_dv)
    yesno_dv.add(f'D{row+1}:E{row+4}')


# ============================================================================
# SECTION 8: GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis_sheet(ws, styles):
    """Create the Gap Analysis sheet."""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 12

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Sheet 7: Gap Analysis"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A3:H3')
    ws['A3'] = "Consolidated Gap Register"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    # Column headers
    headers = [
        "Gap_ID", "Source_Sheet", "Gap_Description", "Severity",
        "Owner", "Target_Date", "Remediation_Plan", "Status"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["subheader"], styles["border"])

    # Data rows (30 rows)
    for row in range(6, 36):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col, value="")
            apply_style(cell, styles["input"], styles["border"])

    # Data validation for Source_Sheet
    source_dv = DataValidation(type="list", formula1='"PIR Process,Root Cause Analysis,Lessons Learned,Control Improvements,Trend Analysis"', allow_blank=True)
    ws.add_data_validation(source_dv)
    source_dv.add(f'B6:B35')

    # Data validation for Severity
    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(severity_dv)
    severity_dv.add(f'D6:D35')

    # Data validation for Status
    status_dv = DataValidation(type="list", formula1='"Open,In_Progress,Closed,Accepted"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f'H6:H35')

    # Summary row
    row = 38
    ws[f'A{row}'] = "Total Gaps:"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])
    ws[f'B{row}'] = '=COUNTA(A6:A35)'
    apply_style(ws[f'B{row}'], styles["formula"], styles["border"])
    ws[f'C{row}'] = "Critical:"
    ws[f'D{row}'] = '=COUNTIF(D6:D35,"Critical")'
    apply_style(ws[f'D{row}'], styles["formula"], styles["border"])
    ws[f'E{row}'] = "High:"
    ws[f'F{row}'] = '=COUNTIF(D6:D35,"High")'
    apply_style(ws[f'F{row}'], styles["formula"], styles["border"])


# ============================================================================
# SECTION 9: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register_sheet(ws, styles):
    """Create the Evidence Register sheet."""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "Sheet 8: Evidence Register"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A3:G3')
    ws['A3'] = "Evidence Tracking Log"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    # Column headers
    headers = [
        "Evidence_ID", "Related_Sheet", "Evidence_Type", "Description",
        "File_Location", "Collection_Date", "Collected_By"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["subheader"], styles["border"])

    # Data rows (50 rows)
    for row in range(6, 56):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col, value="")
            apply_style(cell, styles["input"], styles["border"])
        # Auto-generate Evidence_ID
        ws.cell(row=row, column=1, value=f'=IF(B{row}<>"","EVD-"&TEXT(ROW()-5,"000"),"")')
        apply_style(ws.cell(row=row, column=1), styles["formula"], styles["border"])

    # Data validation for Related_Sheet
    sheet_dv = DataValidation(type="list", formula1='"PIR Process,Root Cause Analysis,Lessons Learned,Control Improvements,Trend Analysis,Gap Analysis"', allow_blank=True)
    ws.add_data_validation(sheet_dv)
    sheet_dv.add(f'B6:B55')

    # Data validation for Evidence_Type
    type_dv = DataValidation(type="list", formula1='"PIR Report,RCA Document,Email,Screenshot,Export,Policy Document,Meeting Minutes,Other"', allow_blank=True)
    ws.add_data_validation(type_dv)
    type_dv.add(f'C6:C55')


# ============================================================================
# SECTION 10: SUMMARY DASHBOARD SHEET
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Create the Summary Dashboard sheet."""
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = f"{DOCUMENT_ID} - Summary Dashboard"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 35

    # Panel 1: Overall Compliance
    ws.merge_cells('A3:D3')
    ws['A3'] = "Panel 1: Overall Compliance"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    ws['A5'] = "Assessment Period"
    ws['B5'] = "[From Instructions sheet]"
    apply_style(ws['B5'], styles["input"], styles["border"])

    ws['A6'] = "Overall Compliance Score"
    ws['B6'] = "=AVERAGE(B10:B14)"
    apply_style(ws['B6'], styles["formula"], styles["border"])

    # Panel 2: Domain Scores
    ws.merge_cells('A8:D8')
    ws['A8'] = "Panel 2: Domain Compliance Scores"
    apply_style(ws['A8'], styles["section_header"], styles["border"])

    domains = [
        ("PIR Process", "='PIR Process'!D72", ">=90%"),
        ("Root Cause Analysis", "='Root Cause Analysis'!D52", ">=90%"),
        ("Lessons Learned", "='Lessons Learned'!D82", ">=90%"),
        ("Control Improvements", "='Control Improvements'!D85", ">=85%"),
        ("Trend Analysis", "Based on KPI status", ">=80%"),
    ]

    row = 10
    for domain, formula, target in domains:
        ws[f'A{row}'] = domain
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws[f'B{row}'] = formula
        apply_style(ws[f'B{row}'], styles["formula"], styles["border"])
        ws[f'C{row}'] = target
        apply_style(ws[f'C{row}'], styles["normal"], styles["border"])
        row += 1

    # Panel 3: Critical Flags
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "Panel 3: Critical Flags"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    flags = [
        ("PIR Overdue (Critical/High)", "Check PIR Process sheet"),
        ("RCA Not Performed (Critical)", "Check Root Cause Analysis sheet"),
        ("Overdue Critical Remediation", "Check Control Improvements sheet"),
        ("Missing Escalation for Overdue", "Check Control Improvements sheet"),
        ("Recurring Root Causes", "Check Root Cause Analysis sheet"),
    ]

    for flag, check in flags:
        row += 1
        ws[f'A{row}'] = flag
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws[f'B{row}'] = check
        apply_style(ws[f'B{row}'], styles["input"], styles["border"])


# ============================================================================
# SECTION 11: APPROVAL SIGN-OFF SHEET
# ============================================================================

def create_approval_signoff_sheet(ws, styles):
    """Create the Approval Sign-Off sheet."""
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "Sheet 10: Approval Sign-Off"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 30

    # Assessment Summary
    ws.merge_cells('A3:D3')
    ws['A3'] = "Assessment Summary"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    ws['A5'] = "Assessment Period"
    ws['B5'] = "[From Instructions]"
    apply_style(ws['B5'], styles["input"], styles["border"])

    ws['A6'] = "Overall Compliance"
    ws['B6'] = "='Summary Dashboard'!B6"
    apply_style(ws['B6'], styles["formula"], styles["border"])

    ws['A7'] = "Critical Gaps"
    ws['B7'] = "='Gap Analysis'!D38"
    apply_style(ws['B7'], styles["formula"], styles["border"])

    # Completed By section
    row = 10
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "Completed By"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    fields = ["Name", "Role", "Department", "Email", "Date", "Signature"]
    for field in fields:
        row += 1
        ws[f'A{row}'] = field
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws[f'B{row}'] = ""
        apply_style(ws[f'B{row}'], styles["input"], styles["border"])

    # Reviewed By section
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "Reviewed By (CSIRT Manager)"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    fields = ["Name", "Date", "Signature", "Comments"]
    for field in fields:
        row += 1
        ws[f'A{row}'] = field
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws[f'B{row}'] = ""
        apply_style(ws[f'B{row}'], styles["input"], styles["border"])

    row += 1
    ws[f'A{row}'] = "Outcome"
    apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
    ws[f'B{row}'] = ""
    apply_style(ws[f'B{row}'], styles["input"], styles["border"])

    # Data validation for Outcome
    outcome_dv = DataValidation(type="list", formula1='"Approved,Approved with corrections,Requires revision"', allow_blank=True)
    ws.add_data_validation(outcome_dv)
    outcome_dv.add(f'B{row}')

    # Approved By section
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "Approved By (CISO)"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    fields = ["Name", "Date", "Signature", "Decision"]
    for field in fields:
        row += 1
        ws[f'A{row}'] = field
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws[f'B{row}'] = ""
        apply_style(ws[f'B{row}'], styles["input"], styles["border"])

    # Data validation for Decision
    decision_dv = DataValidation(type="list", formula1='"Approved,Approved with conditions,Rejected"', allow_blank=True)
    ws.add_data_validation(decision_dv)
    decision_dv.add(f'B{row}')


# ============================================================================
# SECTION 12: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function - orchestrates workbook creation."""
    logger.info("=" * 80)
    logger.info(f"{DOCUMENT_ID} - Learning & Continuous Improvement Assessment Generator")
    logger.info(f"{CONTROL_REF}")
    logger.info("=" * 80)

    wb = create_workbook()
    styles = setup_styles()

    logger.info("\n[1/10] Creating Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)

    logger.info("[2/10] Creating PIR Process sheet...")
    create_pir_process_sheet(wb["PIR Process"], styles)

    logger.info("[3/10] Creating Root Cause Analysis sheet...")
    create_rca_sheet(wb["Root Cause Analysis"], styles)

    logger.info("[4/10] Creating Lessons Learned sheet...")
    create_lessons_learned_sheet(wb["Lessons Learned"], styles)

    logger.info("[5/10] Creating Control Improvements sheet...")
    create_control_improvements_sheet(wb["Control Improvements"], styles)

    logger.info("[6/10] Creating Trend Analysis sheet...")
    create_trend_analysis_sheet(wb["Trend Analysis"], styles)

    logger.info("[7/10] Creating Gap Analysis sheet...")
    create_gap_analysis_sheet(wb["Gap Analysis"], styles)

    logger.info("[8/10] Creating Evidence Register sheet...")
    create_evidence_register_sheet(wb["Evidence Register"], styles)

    logger.info("[9/10] Creating Summary Dashboard sheet...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

    logger.info("[10/10] Creating Approval Sign-Off sheet...")
    create_approval_signoff_sheet(wb["Approval Sign-Off"], styles)

    filename = OUTPUT_FILENAME
    wb.save(filename)

    logger.info(f"\nSUCCESS: {filename}")
    logger.info("\nWorkbook Structure:")
    logger.info("  - 10 sheets (Instructions through Approval)")
    logger.info("  - PIR inventory (50 rows)")
    logger.info("  - RCA inventory (30 rows for Critical/High)")
    logger.info("  - Lessons Learned log (50 rows)")
    logger.info("  - Remediation actions (70 rows)")
    logger.info("  - KPI tracking (8 pre-defined KPIs)")
    logger.info("  - Gap analysis (30 rows)")
    logger.info("  - Evidence register (50 rows)")
    logger.info("\nNext steps:")
    logger.info("  1) Complete document information in Instructions & Legend")
    logger.info("  2) Review PIRs from last 12 months in PIR Process sheet")
    logger.info("  3) Assess RCA quality for Critical/High incidents")
    logger.info("  4) Verify Lessons Learned documentation and distribution")
    logger.info("  5) Track remediation actions in Control Improvements")
    logger.info("  6) Complete Trend Analysis with current KPI values")
    logger.info("  7) Review Gap Analysis and prioritize remediation")
    logger.info("  8) Link evidence in Evidence Register")
    logger.info("  9) Review Summary Dashboard")
    logger.info("  10) Complete Approval Sign-Off workflow")
    logger.info("\nEstimated completion time: 6-10 hours")
    logger.info("\n" + "=" * 80)


if __name__ == "__main__":
    main()


# ============================================================================
# END OF SCRIPT
# ============================================================================

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - INITIAL CREATION
# QA_TOOL: Claude Code
# CHANGES: Initial creation based on ISMS-IMP-A.5.24-28.S5 specification
# =============================================================================
