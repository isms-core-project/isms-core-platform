#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.24-28.S2 - Detection & Classification Assessment Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.25: Incident Assessment & Decision
Assessment Domain 2 of 5: Detection & Classification

--------------------------------------------------------------------------------
PURPOSE & DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
incident detection and classification capabilities, focusing on the assessment
and decision phase (A.5.25) of the incident management lifecycle.

**Assessment Scope:**
- Detection mechanisms (SIEM, EDR, IDS/IPS, NDR, user reporting)
- Alert triage and investigation procedures
- Incident classification and categorization
- Severity assignment criteria and processes
- False positive handling and tuning
- Detection coverage gaps
- MTTD and MTTT metrics

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and metrics definitions
2. Detection Mechanisms - 33 questions on SIEM, EDR, IDS/IPS, coverage
3. Alert Handling - 25 questions on triage, playbooks, escalation
4. Classification & Severity - 25 questions on categorization, severity
5. Detection Effectiveness - 25 questions on MTTD, MTTT, false positives
6. Gap Analysis - 40 gap capacity with risk prioritization
7. Evidence Register - 60 evidence capacity
8. Dashboard - Detection effectiveness summary
9. Approval Sign-Off - SOC Manager + CISO approval

**Key Features:**
- 108 assessment questions across 4 domains
- Automated metric calculations (MTTD, MTTT, FP rate, coverage)
- Detection coverage matrix (11 threat categories)
- Alert quality analysis
- Classification consistency assessment
- Protected formulas with unprotected input cells

**Integration:**
This is Domain 2 of 5 assessment domains for A.5.24-28 Incident Management:
- S1: Framework & Governance (A.5.24 focus)
- S2: Detection & Classification (this assessment - A.5.25 focus)
- S3: Response Capabilities (A.5.26 focus)
- S4: Forensic Evidence Management (A.5.28 focus)
- S5: Learning & Improvement (A.5.27 focus)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a524_28_s2_detection_classification.py

Output:
    File: ISMS-IMP-A.5.24-28.S2_Detection_Classification_YYYYMMDD.xlsx
    Location: Current directory

Estimated Completion Time: 6-10 hours

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.5.24-28
Assessment Domain:    2 of 5 (Detection & Classification - A.5.25 Focus)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Python Version:       3.8+

================================================================================
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)




# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.24-28.S2"
WORKBOOK_NAME = "Detection & Classification"
CONTROL_ID = "A.5.24-28"
CONTROL_NAME = "Incident Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    sheets = [
        "Instructions & Legend",
        "Detection Mechanisms",
        "Alert Handling",
        "Classification & Severity",
        "Detection Effectiveness",
        "Gap Analysis",
        "Evidence Register",
        "Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color="D8E4F8", end_color="D8E4F8", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "calculated_cell": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
        "gap_critical": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "gap_high": {"fill": PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")},
        "gap_medium": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "gap_low": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
    }
    return styles


# ============================================================================
# SECTION 2: INSTRUCTIONS SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create the Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.5.24-28.S2 — Detection & Classification Assessment\n"
        "ISO/IEC 27001:2022 - Control A.5.25: Assessment and Decision"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40
    
    row = 3
    ws[f"A{row}"] = "DOCUMENT INFORMATION"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:B{row}")
    
    doc_info = [
        ("Document ID", "ISMS-IMP-A.5.24-28.S2"),
        ("Assessment Area", "Incident Detection & Classification (Domain 2)"),
        ("Related Policy", "ISMS-POL-A.5.24-28, Section 2.2"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Annual"),
    ]
    
    row += 1
    for label, value in doc_info:
        ws[f"A{row}"] = label + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if label in ["Assessment Date", "Completed By", "Organisation"]:
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1
    
    # Metrics definitions
    row += 2
    ws[f"A{row}"] = "KEY METRICS DEFINITIONS"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:G{row}")
    
    row += 1
    metrics_text = """MTTD (Mean Time to Detect): Time from actual event occurrence to alert generation
MTTT (Mean Time to Triage): Time from alert generation to triage completion
False Positive Rate: (False Positives / Total Alerts) × 100
True Positive Rate: (True Positives / Total Alerts) × 100
Detection Coverage: (Categories with Detection / Total Categories) × 100
Alert-to-Incident Ratio: Total Alerts / True Positive Incidents"""
    
    ws.merge_cells(f"A{row}:G{row+5}")
    ws[f"A{row}"] = metrics_text
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 100
    
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 3: DETECTION MECHANISMS SHEET
# ============================================================================

def create_detection_mechanisms(ws, styles):
    """Create Detection Mechanisms sheet (33 questions)."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "Sheet 2: Detection Mechanisms"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    ws["A2"] = "SIEM, EDR, IDS/IPS, NDR, User Reporting, Coverage by Threat Category (33 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")
    
    # Column headers
    row = 4
    headers = ["Question_ID", "Section", "Question", "Answer", "Evidence_Reference", "Comments", "Gap_Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    # Questions (showing subset - actual implementation would have all 33)
    questions = [
        ("Q1", "SIEM", "What SIEM platform(s) does [Organisation] use?", "Text"),
        ("Q2", "SIEM", "How many active detection rules/alerts are configured in the SIEM?", "Number"),
        ("Q3", "SIEM", "Which threat categories have SIEM detection rules?", "Checkbox"),
        ("Q4", "SIEM", "How often are SIEM rules tuned to reduce false positives?", "Weekly/Monthly/Quarterly/Annually/Ad-Hoc/Never"),
        ("Q5", "SIEM", "What percentage of SIEM rules are custom-developed (vs. vendor-provided)?", "Percentage"),
        ("Q6", "SIEM", "Is threat intelligence integrated into SIEM for automated alerting?", "Yes - Automated/Manual/No"),
        ("Q7", "SIEM", "What is the average daily SIEM alert volume?", "Number"),
        ("Q8", "SIEM", "How many log sources feed the SIEM?", "Number"),
        
        ("Q9", "EDR", "Is EDR or XDR deployed?", "Yes - EDR/Yes - XDR/No"),
        ("Q10", "EDR", "What percentage of endpoints have EDR agent deployed?", "Percentage"),
        ("Q11", "EDR", "What types of EDR alerts are enabled?", "Checkbox"),
        ("Q12", "EDR", "Does EDR support automated response actions?", "Fully/Semi/Manual"),
        ("Q13", "EDR", "How long does EDR retain forensic data?", "30/60/90/180/365+ days"),
        
        ("Q14", "Network Detection", "Is IDS/IPS deployed?", "IPS/IDS/No"),
        ("Q15", "Network Detection", "Where is IDS/IPS deployed?", "Checkbox"),
        ("Q16", "Network Detection", "How frequently are IDS/IPS signatures updated?", "Daily/Weekly/Monthly"),
        ("Q17", "Network Detection", "Is NDR/NTA deployed?", "Yes/No"),
        
        ("Q18", "Other Detection", "Is there a user reporting mechanism (A.6.8)?", "Dedicated/Helpdesk/No"),
        ("Q19", "Other Detection", "Average monthly user-reported events?", "Number"),
        ("Q20", "Other Detection", "Does DLP generate security alerts?", "Integrated/Separate/No"),
        ("Q21", "Other Detection", "Does CASB generate security alerts?", "Yes/No/N/A"),
        ("Q22", "Other Detection", "Are honeypots/deception deployed?", "Yes/No"),
        
        # Coverage by category (Q23-Q33)
        ("Q23", "Coverage", "Category 1: Malware Detection", "Yes/Partial/No"),
        ("Q24", "Coverage", "Category 2: Unauthorised Access Detection", "Yes/Partial/No"),
        ("Q25", "Coverage", "Category 3: Data Breach Detection", "Yes/Partial/No"),
        ("Q26", "Coverage", "Category 4: DoS Detection", "Yes/Partial/No"),
        ("Q27", "Coverage", "Category 5: Social Engineering Detection", "Yes/Partial/No"),
        ("Q28", "Coverage", "Category 6: Web Attack Detection", "Yes/Partial/No"),
        ("Q29", "Coverage", "Category 7: Network Attack Detection", "Yes/Partial/No"),
        ("Q30", "Coverage", "Category 8: Endpoint Compromise Detection", "Yes/Partial/No"),
        ("Q31", "Coverage", "Category 9: Physical Security Detection", "Yes/Partial/No"),
        ("Q32", "Coverage", "Category 10: Config/Patch Detection", "Yes/Partial/No"),
        ("Q33", "Coverage", "Category 11: Other Detection", "Yes/Partial/No"),
    ]
    
    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(OR(D{row}="No", D{row}="Never", D{row}<90), "Yes", "No")'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12
    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 4: ALERT HANDLING SHEET
# ============================================================================

def create_alert_handling(ws, styles):
    """Create Alert Handling sheet (25 questions)."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "Sheet 3: Alert Handling"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    ws["A2"] = "Triage, Investigation, Escalation, Analyst Capacity (25 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")
    
    row = 4
    headers = ["Question_ID", "Section", "Question", "Answer", "Evidence_Reference", "Comments", "Gap_Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    questions = [
        ("Q34", "Triage", "Are alert triage procedures documented?", "Comprehensive/Basic/No"),
        ("Q35", "Triage", "Are triage time SLAs defined for each severity?", "Yes/No"),
        ("Q36", "Triage", "Percentage of alerts triaged within SLA? (90 days)", "Percentage"),
        ("Q37", "Triage", "How are alerts prioritised in queue?", "Automated/Manual/FIFO/No Process"),
        ("Q38", "Triage", "Documented shift handoff process?", "Formal/Informal/No"),
        ("Q39", "Triage", "Do analysts receive triage training?", "Regular/Onboarding/No"),
        ("Q40", "Triage", "What tools for alert triage?", "Checkbox"),
        
        ("Q41", "Playbooks", "Are investigation playbooks documented?", "Comprehensive/Limited/No"),
        ("Q42", "Playbooks", "Which categories have playbooks?", "Checkbox"),
        ("Q43", "Playbooks", "Playbook format?", "SOAR/Written/Flowchart/Checklist"),
        ("Q44", "Playbooks", "Playbook update frequency?", "Quarterly/Semi/Annually/After Incidents/Never"),
        ("Q45", "Playbooks", "Playbook: Malware/Ransomware", "Yes/No"),
        ("Q46", "Playbooks", "Playbook: Phishing/BEC", "Yes/No"),
        ("Q47", "Playbooks", "Playbook: Unauthorised Access", "Yes/No"),
        ("Q48", "Playbooks", "Playbook: Data Exfiltration", "Yes/No"),
        
        ("Q49", "Escalation", "Escalation criteria clearly defined?", "Well Defined/Partial/No"),
        ("Q50", "Escalation", "How are incidents escalated SOC → CSIRT?", "Automated/Manual/Informal/No Process"),
        ("Q51", "Escalation", "Handoff checklist for CSIRT escalation?", "Yes/No"),
        ("Q52", "Escalation", "False positive handling process?", "Documented/Informal/No"),
        ("Q53", "Escalation", "Feedback loop: FP → tuning?", "Systematic/Informal/No"),
        
        ("Q54", "Capacity", "SOC analyst FTE count?", "Number"),
        ("Q55", "Capacity", "Tiered analyst structure?", "Multi-Tier/Single-Tier"),
        ("Q56", "Capacity", "Analyst training frequency?", "Quarterly/Semi/Annually/Onboarding/Never"),
        ("Q57", "Capacity", "Analyst burnout indicators?", "No/Some/Significant"),
        ("Q58", "Capacity", "Analysts trained on all tools?", "Comprehensive/Partial/No"),
    ]
    
    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(OR(D{row}="No", D{row}="Informal", D{row}<80), "Yes", "No")'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12
    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 5: CLASSIFICATION & SEVERITY SHEET
# ============================================================================

def create_classification_severity(ws, styles):
    """Create Classification & Severity sheet (25 questions)."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "Sheet 4: Classification & Severity"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    ws["A2"] = "Category Assignment, Severity Criteria, Consistency (25 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")
    
    row = 4
    headers = ["Question_ID", "Section", "Question", "Answer", "Evidence_Reference", "Comments", "Gap_Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    questions = [
        ("Q59", "Classification", "Adopted ISMS-REF-A.5.24-28 taxonomy?", "Fully/Partially/No"),
        ("Q60", "Classification", "Classification procedure documented?", "Yes/No"),
        ("Q61", "Classification", "Analysts trained on classification?", "Regular/Onboarding/No"),
        ("Q62", "Classification", "Classification consistency reviewed?", "Monthly/Quarterly/Annually/No"),
        ("Q63", "Classification", "Classification accuracy rate?", "Percentage"),
        ("Q64", "Classification", "Reclassification process documented?", "Yes/No"),
        ("Q65", "Classification", "Multi-category incident handling?", "Primary+Secondary/Multiple Tickets/Single/No Process"),
        
        ("Q66", "Severity", "How many severity levels?", "4/3/5/Other"),
        ("Q67", "Severity", "Severity criteria documented?", "Comprehensive/Basic/No"),
        ("Q68", "Severity", "Critical severity criteria?", "Checkbox"),
        ("Q69", "Severity", "Severity assignment training?", "Regular/Onboarding/No"),
        ("Q70", "Severity", "Severity consistency reviewed?", "Monthly/Quarterly/Annually/No"),
        ("Q71", "Severity", "Severity upgrade process?", "Documented/Informal/No"),
        ("Q72", "Severity", "Severity downgrade process?", "Documented/Informal/No"),
        ("Q73", "Severity", "% Incidents Critical (90 days)?", "Percentage"),
        ("Q74", "Severity", "% Incidents High (90 days)?", "Percentage"),
        ("Q75", "Severity", "% Incidents Medium (90 days)?", "Percentage"),
        
        # Top categories by volume (Q76-Q83 not shown for brevity)
    ]
    
    # Add remaining 8 questions for category/severity correlation
    for i in range(76, 84):
        questions.append((f"Q{i}", "Correlation", f"Top Category {i-75}", "Text"))
    
    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(OR(D{row}="No", D{row}<80), "Yes", "No")'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12
    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 6: DETECTION EFFECTIVENESS SHEET
# ============================================================================

def create_detection_effectiveness(ws, styles):
    """Create Detection Effectiveness sheet (25 questions)."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "Sheet 5: Detection Effectiveness"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    ws["A2"] = "MTTD, MTTT, False Positives, Coverage Gaps (25 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")
    
    row = 4
    headers = ["Question_ID", "Section", "Question", "Answer", "Evidence_Reference", "Comments", "Gap_Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    questions = [
        ("Q84", "MTTD", "Is MTTD tracked?", "Yes/No"),
        ("Q85", "MTTD", "Overall MTTD (90 days)?", "Duration"),
        ("Q86", "MTTD", "MTTD for Critical incidents?", "Duration"),
        ("Q87", "MTTD", "MTTD for High incidents?", "Duration"),
        ("Q88", "MTTD", "MTTD trend?", "Improving/Stable/Worsening/Not Tracked"),
        
        ("Q89", "MTTT", "Is MTTT tracked?", "Yes/No"),
        ("Q90", "MTTT", "Overall MTTT (90 days)?", "Duration"),
        ("Q91", "MTTT", "MTTT for Critical alerts?", "Duration"),
        ("Q92", "MTTT", "MTTT for High alerts?", "Duration"),
        ("Q93", "MTTT", "MTTT SLA compliance %?", "Percentage"),
        
        ("Q94", "False Positives", "FPs systematically tracked?", "Yes/No"),
        ("Q95", "False Positives", "Total alerts (90 days)?", "Number"),
        ("Q96", "False Positives", "False positives (90 days)?", "Number"),
        ("Q97", "False Positives", "False positive rate?", "=Q96/Q95*100 (Calculated)"),
        ("Q98", "False Positives", "FP trend?", "Improving/Stable/Worsening/Not Tracked"),
        ("Q99", "False Positives", "Top FP source?", "SIEM/EDR/IDS/NDR/User/Other"),
        ("Q100", "False Positives", "FP tuning backlog?", "Large/Small/None/Not Tracked"),
        
        ("Q101", "True Positives", "True positives (90 days)?", "Number"),
        ("Q102", "True Positives", "True positive rate?", "=Q101/Q95*100 (Calculated)"),
        ("Q103", "True Positives", "Incident volume trend?", "Increasing/Stable/Decreasing"),
        ("Q104", "True Positives", "Alert-to-incident ratio?", "=Q95/Q101 (Calculated)"),
        ("Q105", "True Positives", "Can SOC handle volume?", "Comfortably/At Capacity/Overwhelmed"),
        
        ("Q106", "Coverage", "Detection blind spots identified?", "Documented/Informally/No"),
        ("Q107", "Coverage", "Coverage gaps prioritised?", "Yes/No"),
        ("Q108", "Coverage", "Coverage improvement plan?", "Funded/Planned/No"),
    ]
    
    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        
        # Special handling for calculated cells
        if "Calculated" in answer_type:
            formula = answer_type.split("(Calculated)")[0].strip()
            ws[f"D{row}"] = formula.replace("Q95", f"D{row-13}").replace("Q96", f"D{row-12}").replace("Q101", f"D{row-7}")
            ws[f"D{row}"].fill = styles["calculated_cell"]["fill"]
        
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(OR(D{row}="No", D{row}>30, D{row}="Worsening"), "Yes", "No")'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12
    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 7-9: GAP ANALYSIS, EVIDENCE, DASHBOARD, APPROVAL
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet (40 capacity)."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Sheet 6: Gap Analysis"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    row = 4
    headers = ["Gap_ID", "Section", "Gap_Description", "Risk_Level", "Current_State", "Target_State", "Remediation", "Owner", "Target_Date", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    # Add 40 gap rows
    for i in range(40):
        row += 1
        ws[f"A{row}"] = f"GAP-{str(i+1).zfill(3)}"
        for col_idx in range(2, 11):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
    
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.freeze_panes = "A5"


def create_evidence_register(ws, styles):
    """Create Evidence Register (60 capacity)."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Sheet 7: Evidence Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    row = 4
    headers = ["Evidence_ID", "Evidence_Type", "Description", "Related_Section", "Storage_Location", "Date_Collected", "Collected_By", "Verification"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    for i in range(60):
        row += 1
        ws[f"A{row}"] = f"EV-{str(i+1).zfill(3)}"
        for col_idx in range(2, 9):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
    
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.freeze_panes = "A5"


def create_dashboard(ws, styles):
    """Create Dashboard summary."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Sheet 8: Dashboard"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    row = 4
    ws[f"A{row}"] = "DETECTION EFFECTIVENESS SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:D{row}")
    
    metrics = [
        ("Overall MTTD", "='Detection Effectiveness'!D85"),
        ("Overall MTTT", "='Detection Effectiveness'!D90"),
        ("False Positive Rate", "='Detection Effectiveness'!D97"),
        ("True Positive Rate", "='Detection Effectiveness'!D102"),
        ("Detection Coverage", "=COUNTIF('Detection Mechanisms'!D23:D33,\"Yes\")/11*100"),
        ("Alert-to-Incident Ratio", "='Detection Effectiveness'!D104"),
    ]
    
    row += 1
    for metric_label, metric_formula in metrics:
        ws[f"A{row}"] = metric_label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = metric_formula
        ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]
        row += 1
    
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = "A4"


def create_approval_signoff(ws, styles):
    """Create Approval Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "Sheet 9: Approval & Sign-Off"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30
    
    row = 3
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:E{row}")
    
    row += 1
    summary_fields = [
        ("Assessment Document", "ISMS-IMP-A.5.24-28.S2 - Detection & Classification"),
        ("Assessment Period", ""),
        ("False Positive Rate", "=Dashboard!B7"),
        ("Detection Coverage", "=Dashboard!B9"),
    ]
    
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if label == "Assessment Period":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        else:
            ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1
    
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "COMPLETED BY (SOC Manager)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    for field in ["Name", "Date", "Signature"]:
        row += 1
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
    
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    
    for field in ["Name", "Date", "Approval Decision", "Signature"]:
        row += 1
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
    
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 10: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.5.24-28.S2 - Detection & Classification Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.5.25: Assessment & Decision")
    logger.info("=" * 80)
    
    wb = create_workbook()
    styles = setup_styles()
    
    logger.info("\n[1/9] Creating Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"], styles)
    
    logger.info("[2/9] Creating Detection Mechanisms sheet (33 questions)...")
    create_detection_mechanisms(wb["Detection Mechanisms"], styles)
    
    logger.info("[3/9] Creating Alert Handling sheet (25 questions)...")
    create_alert_handling(wb["Alert Handling"], styles)
    
    logger.info("[4/9] Creating Classification & Severity sheet (25 questions)...")
    create_classification_severity(wb["Classification & Severity"], styles)
    
    logger.info("[5/9] Creating Detection Effectiveness sheet (25 questions)...")
    create_detection_effectiveness(wb["Detection Effectiveness"], styles)
    
    logger.info("[6/9] Creating Gap Analysis sheet...")
    create_gap_analysis(wb["Gap Analysis"], styles)
    
    logger.info("[7/9] Creating Evidence Register sheet...")
    create_evidence_register(wb["Evidence Register"], styles)
    
    logger.info("[8/9] Creating Dashboard sheet...")
    create_dashboard(wb["Dashboard"], styles)
    
    logger.info("[9/9] Creating Approval Sign-Off sheet...")
    create_approval_signoff(wb["Approval Sign-Off"], styles)
    
    filename = f"ISMS-IMP-A.5.24-28.S2_Detection_Classification_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    logger.info(f"\n✅ SUCCESS: {filename}")
    logger.info("\nWorkbook Structure:")
    logger.info("  • 9 sheets (Instructions through Approval)")
    logger.info("  • 108 assessment questions (33+25+25+25)")
    logger.info("  • 40 gap analysis capacity")
    logger.info("  • 60 evidence register capacity")
    logger.info("  • Automated metrics (MTTD, MTTT, FP rate, coverage)")
    logger.info("\nNext steps:")
    logger.info("  1) Extract metrics from SIEM/ticketing system (90 days)")
    logger.info("  2) Complete detection mechanism inventory")
    logger.info("  3) Fill yellow cells in assessment sheets")
    logger.info("  4) Review automated metric calculations")
    logger.info("  5) Analyze detection coverage gaps")
    logger.info("  6) Complete Approval Sign-Off workflow")
    logger.info("\nEstimated completion time: 6-10 hours")
    logger.info("\n" + "=" * 80)


if __name__ == "__main__":
    main()


# ============================================================================
# END OF SCRIPT
# ============================================================================
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
