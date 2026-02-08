#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.1-2-6.1-2.S5 - Governance Compliance Dashboard Generator
================================================================================

Purpose:
    Generate Excel governance compliance dashboard for ISO/IEC 27001:2022
    Controls A.5.1, A.5.2, A.6.1, and A.6.2 (Secure Employment and Roles).
    Aggregates findings from S1-S4 workbooks into board-facing executive view.

Control Alignment:
    - ISO/IEC 27001:2022 Annex A.5.1: Policies for Information Security
    - ISO/IEC 27001:2022 Annex A.5.2: Information Security Roles and Responsibilities
    - ISO/IEC 27001:2022 Annex A.6.1: Screening
    - ISO/IEC 27001:2022 Annex A.6.2: Terms and Conditions of Employment
    - Swiss FADP (nFADP) - Federal Act on Data Protection

Assessment Focus:
    - Aggregated S1-S4 findings
    - Lifecycle-weighted scoring
    - Cross-domain gap identification
    - FADP compliance indicator
    - NDA compliance rate KPI
    - CISO certification with FADP acknowledgment

Dashboard Weights:
    - Policy Framework (A.5.1): 35%
    - Roles & Responsibilities (A.5.2): 25%
    - Screening & Vetting (A.6.1): 20%
    - Employment Terms (A.6.2): 20%

Workbook Structure:
    11 sheets total:
    1. Instructions_Legend - Navigation guide, color coding, dropdown reference
    2. Executive_Summary - Board-facing single-page summary with weighted score
    3. Maturity_Assessment - 5-level maturity model + trend
    4. Policy_Summary - S1 (A.5.1) domain metrics
    5. Roles_Summary - S2 (A.5.2) domain metrics
    6. Screening_Summary - S3 (A.6.1) domain metrics + FADP indicator
    7. Contract_Summary - S4 (A.6.2/A.6.5) domain metrics + NDA focal point
    8. Gap_Analysis - Consolidated gaps with cross-domain flags
    9. Evidence_Register - Consolidated evidence index
    10. Action_Items - Remediation tracker
    11. Approval_Sign_Off - Three-level governance certification

Dependencies:
    - openpyxl >= 3.0.0

Usage:
    python generate_a5_1_2_6_1_2_s5_governance_dashboard.py

Output:
    ISMS-IMP-A.5.1-2-6.1-2.S5_Governance_Dashboard_YYYYMMDD.xlsx

Technical Specification:
    See ISMS-IMP-A.5.1-2-6.1-2.S5-Governance-Dashboard.md

Quality Standards:
    - Follows A.8.23/A.8.24 reference implementation patterns
    - Comprehensive documentation and inline comments
    - Consistent styling (Yellow input, Light Blue calculated, Gray labels)
    - Data validation on all appropriate cells
    - Conditional formatting for compliance statuses
    - Sheet protection with selective unlocking
    - Named ranges for cross-sheet references
    - Formula-based auto-calculations
    - Professional formatting for executive presentation

Author:               [Organization] ISMS Implementation Team
Version: 1.0
Date:                 [Date to be set]
License: Internal Use Only

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.1-2-6.1-2.S5"
WORKBOOK_NAME = "Governance Compliance Dashboard"
CONTROL_ID = "A.5.1-2-6.1-2"
CONTROL_NAME = "Information Security Policies and Organization"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# Dashboard Weights
WEIGHT_POLICY = 0.35      # A.5.1 Policy Framework
WEIGHT_ROLES = 0.25       # A.5.2 Roles & Responsibilities
WEIGHT_SCREENING = 0.20   # A.6.1 Screening & Vetting
WEIGHT_EMPLOYMENT = 0.20  # A.6.2 Employment Terms


# ==============================================================================
# STYLE DEFINITIONS
# ==============================================================================

def setup_styles():
    """
    Define consistent cell styles used throughout the workbook.

    Returns:
        dict: Dictionary of style objects for different cell types
    """
    styles = {}

    # Header styles
    styles['header_main'] = {
        'font': Font(name='Calibri', size=18, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['header_sub'] = {
        'font': Font(name='Calibri', size=12, bold=False, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['section_header'] = {
        'font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['subsection_header'] = {
        'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['column_header'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    # Data cell styles
    styles['input_cell'] = {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    styles['calculated_cell'] = {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    styles['label_cell'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    styles['kpi_value'] = {
        'font': Font(name='Calibri', size=16, bold=True, color='000000'),
        'fill': PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    # Status fills (for conditional formatting reference)
    styles['status_compliant'] = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    styles['status_partial'] = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    styles['status_non_compliant'] = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    styles['status_fadp'] = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')

    return styles


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def apply_cell_style(cell, style_dict):
    """
    Apply a style dictionary to a cell.

    Args:
        cell: openpyxl cell object
        style_dict: Dictionary containing style attributes
    """
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']


def set_column_widths(ws, widths_dict):
    """
    Set column widths for a worksheet.

    Args:
        ws: Worksheet object
        widths_dict: Dictionary mapping column letters to widths
    """
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width


def create_data_validation(formula_or_list, allow_blank=True, error_title="Invalid Entry",
                           error_message="Please select from dropdown"):
    """
    Create a data validation object for dropdown lists.

    Args:
        formula_or_list: Either a string formula or list of values
        allow_blank: Whether to allow blank cells
        error_title: Title for error message
        error_message: Error message text

    Returns:
        DataValidation object
    """
    if isinstance(formula_or_list, list):
        dv = DataValidation(type="list", formula1=f'"{",".join(formula_or_list)}"',
                           allow_blank=allow_blank)
    else:
        dv = DataValidation(type="list", formula1=formula_or_list, allow_blank=allow_blank)

    dv.error = error_message
    dv.errorTitle = error_title
    dv.showErrorMessage = True
    return dv


def merge_and_style(ws, range_string, value, style_dict):
    """
    Merge cells and apply style.

    Args:
        ws: Worksheet object
        range_string: Range to merge (e.g., 'A1:K1')
        value: Value to write
        style_dict: Style dictionary to apply
    """
    ws.merge_cells(range_string)
    cell = ws[range_string.split(':')[0]]
    cell.value = value
    apply_cell_style(cell, style_dict)


# ==============================================================================
# SHEET CREATION FUNCTIONS
# ==============================================================================

def create_instructions_legend_sheet(wb, styles):
    """
    Sheet 1: Instructions_Legend

    Navigation guide, color coding, dropdown reference.
    """
    ws = wb.active
    ws.title = "Instructions_Legend"

    # --- Header Section ---
    merge_and_style(ws, 'A1:H1',
                   f'{DOCUMENT_ID} - Governance Compliance Dashboard',
                   styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:H2',
                   'Instructions, Navigation Guide, and Legend',
                   styles['header_sub'])
    ws.row_dimensions[2].height = 30

    # --- Purpose Section ---
    merge_and_style(ws, 'A4:H4', 'PURPOSE', styles['section_header'])

    purpose_text = (
        "This workbook provides an integrated governance compliance dashboard for the "
        "Secure Employment and Roles framework (A.5.1, A.5.2, A.6.1, A.6.2). It aggregates "
        "findings from S1-S4 assessment workbooks into a board-facing executive summary "
        "with weighted compliance scoring and cross-domain gap identification."
    )
    ws['A5'] = purpose_text
    ws.merge_cells('A5:H7')
    ws['A5'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[5].height = 60

    # --- Workbook Structure ---
    merge_and_style(ws, 'A9:H9', 'WORKBOOK STRUCTURE', styles['section_header'])

    sheets = [
        ('1', 'Instructions_Legend', 'Navigation guide and legend (this sheet)'),
        ('2', 'Executive_Summary', 'Board-facing summary with weighted compliance score'),
        ('3', 'Maturity_Assessment', '5-level maturity model with trend analysis'),
        ('4', 'Policy_Summary', 'S1 (A.5.1) Policy Framework domain metrics'),
        ('5', 'Roles_Summary', 'S2 (A.5.2) Roles & Responsibilities domain metrics'),
        ('6', 'Screening_Summary', 'S3 (A.6.1) Screening domain metrics + FADP indicator'),
        ('7', 'Contract_Summary', 'S4 (A.6.2/A.6.5) Employment terms + NDA focal point'),
        ('8', 'Gap_Analysis', 'Consolidated gaps with cross-domain flags'),
        ('9', 'Evidence_Register', 'Consolidated evidence index'),
        ('10', 'Action_Items', 'Remediation tracker'),
        ('11', 'Approval_Sign_Off', 'Three-level governance certification with FADP')
    ]

    for col, header in zip(['A', 'B', 'C'], ['#', 'Sheet Name', 'Description']):
        ws[f'{col}11'] = header
        apply_cell_style(ws[f'{col}11'], styles['column_header'])

    for i, (num, name, desc) in enumerate(sheets, start=12):
        ws[f'A{i}'] = num
        ws[f'B{i}'] = name
        ws[f'C{i}'] = desc
        for col in ['A', 'B', 'C']:
            ws[f'{col}{i}'].border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

    # --- Color Legend ---
    merge_and_style(ws, 'A25:H25', 'COLOR LEGEND', styles['section_header'])

    colors = [
        ('Yellow', 'FFFFCC', 'User Input - Cells requiring manual entry'),
        ('Light Blue', 'DCE6F1', 'Calculated - Auto-populated/formula cells'),
        ('Gray', 'D9D9D9', 'Labels - Row/column headers'),
        ('Green', 'C6EFCE', 'Compliant - Meets requirements'),
        ('Amber/Yellow', 'FFEB9C', 'Partial - Partially meets requirements'),
        ('Red', 'FFC7CE', 'Non-Compliant - Does not meet requirements'),
        ('Purple', '7030A0', 'FADP Relevant - Swiss data protection applicable')
    ]

    for col, header in zip(['A', 'B', 'C'], ['Color', 'Sample', 'Meaning']):
        ws[f'{col}27'] = header
        apply_cell_style(ws[f'{col}27'], styles['column_header'])

    for i, (name, hex_color, meaning) in enumerate(colors, start=28):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = ''
        ws[f'B{i}'].fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
        ws[f'C{i}'] = meaning
        for col in ['A', 'B', 'C']:
            ws[f'{col}{i}'].border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

    # --- Weighting Formula ---
    merge_and_style(ws, 'A37:H37', 'COMPLIANCE WEIGHTING', styles['section_header'])

    weights = [
        ('Policy Framework (A.5.1)', '35%', 'S1 Workbook'),
        ('Roles & Responsibilities (A.5.2)', '25%', 'S2 Workbook'),
        ('Screening & Vetting (A.6.1)', '20%', 'S3 Workbook'),
        ('Employment Terms (A.6.2)', '20%', 'S4 Workbook'),
        ('TOTAL', '100%', '-')
    ]

    for col, header in zip(['A', 'B', 'C'], ['Domain', 'Weight', 'Source']):
        ws[f'{col}39'] = header
        apply_cell_style(ws[f'{col}39'], styles['column_header'])

    for i, (domain, weight, source) in enumerate(weights, start=40):
        ws[f'A{i}'] = domain
        ws[f'B{i}'] = weight
        ws[f'C{i}'] = source
        for col in ['A', 'B', 'C']:
            ws[f'{col}{i}'].border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        if domain == 'TOTAL':
            for col in ['A', 'B', 'C']:
                ws[f'{col}{i}'].font = Font(bold=True)

    # --- Key Indicators ---
    merge_and_style(ws, 'A47:H47', 'KEY INDICATORS', styles['section_header'])

    indicators = [
        ('FADP Indicator', 'Swiss Federal Act on Data Protection compliance flag'),
        ('NDA Rate', 'Percentage of employees with signed NDAs'),
        ('Cross-Domain Gap', 'Gap affecting multiple control domains'),
        ('Maturity Level', 'CMM-style 1-5 maturity assessment')
    ]

    for col, header in zip(['A', 'B'], ['Indicator', 'Description']):
        ws[f'{col}49'] = header
        apply_cell_style(ws[f'{col}49'], styles['column_header'])

    for i, (indicator, desc) in enumerate(indicators, start=50):
        ws[f'A{i}'] = indicator
        ws[f'B{i}'] = desc
        for col in ['A', 'B']:
            ws[f'{col}{i}'].border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

    # Set column widths
    set_column_widths(ws, {'A': 30, 'B': 25, 'C': 50, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15})

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False


def create_executive_summary_sheet(wb, styles):
    """
    Sheet 2: Executive_Summary

    Board-facing single-page summary with weighted compliance score.
    """
    ws = wb.create_sheet("Executive_Summary")

    # --- Header ---
    merge_and_style(ws, 'A1:J1',
                   f'{DOCUMENT_ID} - Executive Summary',
                   styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:J2',
                   'Governance Compliance Dashboard - Board-Level View',
                   styles['header_sub'])
    ws.row_dimensions[2].height = 30

    # --- Document Information Block ---
    info_items = [
        (4, 'Document ID:', DOCUMENT_ID),
        (5, 'Control Reference:', CONTROL_REF),
        (6, 'Assessment Period:', None),
        (7, 'Assessment Date:', '=TODAY()'),
        (8, 'Prepared By:', None),
        (9, 'Framework Version:', '1.0')
    ]

    for row_num, label, value in info_items:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        ws.merge_cells(f'B{row_num}:C{row_num}')

        if value:
            ws[f'B{row_num}'] = value
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # --- Overall Compliance Score (Large KPI) ---
    merge_and_style(ws, 'E4:G5', 'OVERALL COMPLIANCE', styles['section_header'])

    ws.merge_cells('E6:G8')
    ws['E6'] = '=ROUND((Policy_Summary!B20*0.35)+(Roles_Summary!B20*0.25)+(Screening_Summary!B20*0.20)+(Contract_Summary!B20*0.20),1)'
    ws['E6'].number_format = '0.0%'
    apply_cell_style(ws['E6'], styles['kpi_value'])
    ws['E6'].font = Font(name='Calibri', size=28, bold=True)
    ws.row_dimensions[6].height = 30
    ws.row_dimensions[7].height = 30
    ws.row_dimensions[8].height = 30

    # --- FADP Compliance Indicator ---
    merge_and_style(ws, 'I4:J5', 'FADP STATUS', styles['section_header'])
    ws['I4'].fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')

    ws.merge_cells('I6:J8')
    ws['I6'] = '=IF(Screening_Summary!B22="Yes","COMPLIANT",IF(Screening_Summary!B22="Partial","PARTIAL","REVIEW REQUIRED"))'
    apply_cell_style(ws['I6'], styles['kpi_value'])
    ws['I6'].font = Font(name='Calibri', size=14, bold=True)

    # --- Domain Breakdown ---
    merge_and_style(ws, 'A11:J11', 'DOMAIN COMPLIANCE BREAKDOWN', styles['section_header'])

    headers = ['Domain', 'Weight', 'Raw Score', 'Weighted Score', 'Status', 'Source Workbook']
    for i, header in enumerate(headers, start=1):
        col = get_column_letter(i)
        ws[f'{col}13'] = header
        apply_cell_style(ws[f'{col}13'], styles['column_header'])

    domains = [
        ('Policy Framework (A.5.1)', '35%', '=Policy_Summary!B20', '=C14*0.35', 'S1'),
        ('Roles & Responsibilities (A.5.2)', '25%', '=Roles_Summary!B20', '=C15*0.25', 'S2'),
        ('Screening & Vetting (A.6.1)', '20%', '=Screening_Summary!B20', '=C16*0.20', 'S3'),
        ('Employment Terms (A.6.2)', '20%', '=Contract_Summary!B20', '=C17*0.20', 'S4')
    ]

    for i, (domain, weight, raw_formula, weighted_formula, source) in enumerate(domains, start=14):
        ws[f'A{i}'] = domain
        apply_cell_style(ws[f'A{i}'], styles['label_cell'])

        ws[f'B{i}'] = weight
        apply_cell_style(ws[f'B{i}'], styles['calculated_cell'])

        ws[f'C{i}'] = raw_formula
        ws[f'C{i}'].number_format = '0.0%'
        apply_cell_style(ws[f'C{i}'], styles['calculated_cell'])

        ws[f'D{i}'] = weighted_formula
        ws[f'D{i}'].number_format = '0.0%'
        apply_cell_style(ws[f'D{i}'], styles['calculated_cell'])

        ws[f'E{i}'] = f'=IF(C{i}>=0.9,"Compliant",IF(C{i}>=0.7,"Partial","Non-Compliant"))'
        apply_cell_style(ws[f'E{i}'], styles['calculated_cell'])

        ws[f'F{i}'] = source
        apply_cell_style(ws[f'F{i}'], styles['calculated_cell'])

    # Total row
    ws['A18'] = 'WEIGHTED TOTAL'
    apply_cell_style(ws['A18'], styles['label_cell'])
    ws['A18'].font = Font(name='Calibri', size=10, bold=True)

    ws['B18'] = '100%'
    ws['D18'] = '=SUM(D14:D17)'
    ws['D18'].number_format = '0.0%'
    ws['D18'].font = Font(name='Calibri', size=10, bold=True)
    apply_cell_style(ws['D18'], styles['calculated_cell'])

    # --- Key KPIs ---
    merge_and_style(ws, 'A20:J20', 'KEY PERFORMANCE INDICATORS', styles['section_header'])

    kpis = [
        ('NDA Compliance Rate', '=Contract_Summary!B24', '>=95%', 'Employment'),
        ('Policy Acknowledgment Rate', '=Policy_Summary!B24', '>=90%', 'Policy'),
        ('Screening Completion Rate', '=Screening_Summary!B24', '100%', 'HR'),
        ('FADP Training Completion', '=Screening_Summary!B26', '>=95%', 'Training'),
        ('Critical Gaps Open', '=Gap_Analysis!B50', '0', 'Risk'),
        ('Cross-Domain Gaps', '=Gap_Analysis!B51', '0', 'Risk')
    ]

    for col, header in zip(['A', 'B', 'C', 'D', 'E'],
                           ['KPI', 'Current Value', 'Target', 'Status', 'Owner']):
        ws[f'{col}22'] = header
        apply_cell_style(ws[f'{col}22'], styles['column_header'])

    for i, (kpi, formula, target, owner) in enumerate(kpis, start=23):
        ws[f'A{i}'] = kpi
        apply_cell_style(ws[f'A{i}'], styles['label_cell'])

        ws[f'B{i}'] = formula
        apply_cell_style(ws[f'B{i}'], styles['calculated_cell'])
        if 'Rate' in kpi or 'Completion' in kpi:
            ws[f'B{i}'].number_format = '0.0%'

        ws[f'C{i}'] = target
        apply_cell_style(ws[f'C{i}'], styles['calculated_cell'])

        ws[f'D{i}'] = f'=IF(B{i}>=VALUE(SUBSTITUTE(C{i},"%","")/100),"On Target","Below Target")'
        apply_cell_style(ws[f'D{i}'], styles['calculated_cell'])

        ws[f'E{i}'] = owner
        apply_cell_style(ws[f'E{i}'], styles['calculated_cell'])

    # --- Assessment Status ---
    merge_and_style(ws, 'A31:J31', 'ASSESSMENT STATUS', styles['section_header'])

    ws['A33'] = 'Overall Assessment Status:'
    apply_cell_style(ws['A33'], styles['label_cell'])
    ws.merge_cells('B33:C33')
    apply_cell_style(ws['B33'], styles['input_cell'])

    status_dv = create_data_validation(['Draft', 'Under-Review', 'Approved', 'Audit-Ready'])
    ws.add_data_validation(status_dv)
    status_dv.add('B33')

    ws['A34'] = 'Next Board Review:'
    apply_cell_style(ws['A34'], styles['label_cell'])
    ws.merge_cells('B34:C34')
    apply_cell_style(ws['B34'], styles['input_cell'])
    ws['B34'].number_format = 'DD.MM.YYYY'

    ws['A35'] = 'CISO Recommendation:'
    apply_cell_style(ws['A35'], styles['label_cell'])
    ws.merge_cells('B35:E35')
    apply_cell_style(ws['B35'], styles['input_cell'])

    # Set column widths
    set_column_widths(ws, {'A': 35, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 20, 'G': 15, 'H': 15, 'I': 15, 'J': 15})

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False


def create_maturity_assessment_sheet(wb, styles):
    """
    Sheet 3: Maturity_Assessment

    5-level maturity model + trend analysis.
    """
    ws = wb.create_sheet("Maturity_Assessment")

    # --- Header ---
    merge_and_style(ws, 'A1:I1',
                   'MATURITY ASSESSMENT',
                   styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:I2',
                   '5-Level Capability Maturity Model for Secure Employment Framework',
                   styles['header_sub'])

    # --- Maturity Level Definitions ---
    merge_and_style(ws, 'A4:I4', 'MATURITY LEVEL DEFINITIONS', styles['section_header'])

    levels = [
        ('Level 1 - Initial', 'Ad-hoc processes, inconsistent execution, reactive approach'),
        ('Level 2 - Repeatable', 'Basic processes documented, some consistency'),
        ('Level 3 - Defined', 'Standardized processes, proactive management'),
        ('Level 4 - Managed', 'Measured processes, quantitative goals'),
        ('Level 5 - Optimized', 'Continuous improvement, best practice leadership')
    ]

    for col, header in zip(['A', 'B'], ['Level', 'Description']):
        ws[f'{col}6'] = header
        apply_cell_style(ws[f'{col}6'], styles['column_header'])

    for i, (level, desc) in enumerate(levels, start=7):
        ws[f'A{i}'] = level
        ws.merge_cells(f'B{i}:I{i}')
        ws[f'B{i}'] = desc
        ws[f'A{i}'].border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))
        ws[f'B{i}'].border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Domain Maturity Assessment ---
    merge_and_style(ws, 'A14:I14', 'DOMAIN MATURITY ASSESSMENT', styles['section_header'])

    headers = ['Domain', 'Control', 'Current Level', 'Target Level', 'Gap', 'Trend', 'Priority', 'Notes']
    for i, header in enumerate(headers, start=1):
        col = get_column_letter(i)
        ws[f'{col}16'] = header
        apply_cell_style(ws[f'{col}16'], styles['column_header'])

    domain_assessments = [
        ('Policy Framework', 'A.5.1'),
        ('Roles & Responsibilities', 'A.5.2'),
        ('Screening & Vetting', 'A.6.1'),
        ('Employment Terms', 'A.6.2')
    ]

    for i, (domain, control) in enumerate(domain_assessments, start=17):
        ws[f'A{i}'] = domain
        apply_cell_style(ws[f'A{i}'], styles['label_cell'])

        ws[f'B{i}'] = control
        apply_cell_style(ws[f'B{i}'], styles['calculated_cell'])

        # Current Level - input
        apply_cell_style(ws[f'C{i}'], styles['input_cell'])

        # Target Level - input
        apply_cell_style(ws[f'D{i}'], styles['input_cell'])

        # Gap - calculated
        ws[f'E{i}'] = f'=IF(AND(C{i}<>"",D{i}<>""),D{i}-C{i},"")'
        apply_cell_style(ws[f'E{i}'], styles['calculated_cell'])

        # Trend - input
        apply_cell_style(ws[f'F{i}'], styles['input_cell'])

        # Priority - calculated
        ws[f'G{i}'] = f'=IF(E{i}="","",IF(E{i}>=2,"High",IF(E{i}>=1,"Medium","Low")))'
        apply_cell_style(ws[f'G{i}'], styles['calculated_cell'])

        # Notes - input
        apply_cell_style(ws[f'H{i}'], styles['input_cell'])

    # Data validation for maturity levels
    level_dv = create_data_validation(['1', '2', '3', '4', '5'])
    ws.add_data_validation(level_dv)
    level_dv.add('C17:D20')

    trend_dv = create_data_validation(['Improving', 'Stable', 'Declining'])
    ws.add_data_validation(trend_dv)
    trend_dv.add('F17:F20')

    # --- Overall Maturity Summary ---
    merge_and_style(ws, 'A23:I23', 'OVERALL MATURITY SUMMARY', styles['section_header'])

    ws['A25'] = 'Average Current Maturity:'
    apply_cell_style(ws['A25'], styles['label_cell'])
    ws['B25'] = '=AVERAGE(C17:C20)'
    ws['B25'].number_format = '0.0'
    apply_cell_style(ws['B25'], styles['calculated_cell'])

    ws['A26'] = 'Average Target Maturity:'
    apply_cell_style(ws['A26'], styles['label_cell'])
    ws['B26'] = '=AVERAGE(D17:D20)'
    ws['B26'].number_format = '0.0'
    apply_cell_style(ws['B26'], styles['calculated_cell'])

    ws['A27'] = 'Overall Gap:'
    apply_cell_style(ws['A27'], styles['label_cell'])
    ws['B27'] = '=B26-B25'
    ws['B27'].number_format = '0.0'
    apply_cell_style(ws['B27'], styles['calculated_cell'])

    # --- Trend History ---
    merge_and_style(ws, 'A30:I30', 'TREND HISTORY', styles['section_header'])

    trend_headers = ['Assessment Date', 'Policy (A.5.1)', 'Roles (A.5.2)',
                     'Screening (A.6.1)', 'Terms (A.6.2)', 'Average', 'Notes']
    for i, header in enumerate(trend_headers, start=1):
        col = get_column_letter(i)
        ws[f'{col}32'] = header
        apply_cell_style(ws[f'{col}32'], styles['column_header'])

    # Trend history rows (5 historical entries)
    for row in range(33, 38):
        for col in ['A', 'B', 'C', 'D', 'E', 'G']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])
        ws[f'A{row}'].number_format = 'DD.MM.YYYY'
        ws[f'F{row}'] = f'=IF(B{row}<>"",AVERAGE(B{row}:E{row}),"")'
        ws[f'F{row}'].number_format = '0.0'
        apply_cell_style(ws[f'F{row}'], styles['calculated_cell'])

    # Set column widths
    set_column_widths(ws, {'A': 30, 'B': 15, 'C': 15, 'D': 15, 'E': 10, 'F': 15, 'G': 15, 'H': 40, 'I': 15})

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False


def create_policy_summary_sheet(wb, styles):
    """
    Sheet 4: Policy_Summary

    S1 (A.5.1) domain metrics from Policy Framework Assessment.
    """
    ws = wb.create_sheet("Policy_Summary")

    # --- Header ---
    merge_and_style(ws, 'A1:F1',
                   'POLICY FRAMEWORK SUMMARY (A.5.1)',
                   styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:F2',
                   'S1 Workbook Metrics - Policies for Information Security',
                   styles['header_sub'])

    # --- Source Reference ---
    ws['A4'] = 'Source Workbook:'
    apply_cell_style(ws['A4'], styles['label_cell'])
    ws['B4'] = 'ISMS-IMP-A.5.1-2-6.1-2.S1_Policy_Framework'
    apply_cell_style(ws['B4'], styles['calculated_cell'])

    ws['A5'] = 'Weight in Dashboard:'
    apply_cell_style(ws['A5'], styles['label_cell'])
    ws['B5'] = '35%'
    apply_cell_style(ws['B5'], styles['calculated_cell'])

    # --- Key Metrics ---
    merge_and_style(ws, 'A7:F7', 'KEY METRICS', styles['section_header'])

    metrics = [
        (9, 'Total Policies Assessed:', None, True),
        (10, 'Policies Compliant:', None, True),
        (11, 'Policies Partial:', None, True),
        (12, 'Policies Non-Compliant:', None, True),
        (13, 'Tier-1 Policies:', None, True),
        (14, 'Tier-2 Policies:', None, True),
        (15, 'Tier-3 Policies:', None, True),
        (16, 'Average Policy Age (months):', None, True),
        (17, 'Overdue Reviews:', None, True),
        (18, 'Policies with Gaps:', None, True)
    ]

    for row_num, label, value, is_input in metrics:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        if is_input:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])
        else:
            ws[f'B{row_num}'] = value
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])

    # --- Overall Compliance Score ---
    merge_and_style(ws, 'A20:B20', 'OVERALL COMPLIANCE SCORE', styles['subsection_header'])
    ws['A20'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')

    # This would normally be calculated from S1 workbook data
    # For dashboard purposes, it's an input that should match S1 dashboard
    apply_cell_style(ws['B20'], styles['input_cell'])
    ws['B20'].number_format = '0.0%'

    # --- Compliance Breakdown ---
    merge_and_style(ws, 'A22:F22', 'COMPLIANCE BY DOMAIN', styles['section_header'])

    domains = [
        ('Policy Inventory Completeness', '25%'),
        ('Policy Lifecycle Compliance', '25%'),
        ('Policy Governance', '20%'),
        ('Policy Classification', '10%'),
        ('Policy Communication', '15%'),
        ('Policy Repository', '5%')
    ]

    for col, header in zip(['A', 'B', 'C', 'D'], ['Domain', 'Weight', 'Score', 'Status']):
        ws[f'{col}24'] = header
        apply_cell_style(ws[f'{col}24'], styles['column_header'])

    for i, (domain, weight) in enumerate(domains, start=25):
        ws[f'A{i}'] = domain
        apply_cell_style(ws[f'A{i}'], styles['label_cell'])

        ws[f'B{i}'] = weight
        apply_cell_style(ws[f'B{i}'], styles['calculated_cell'])

        apply_cell_style(ws[f'C{i}'], styles['input_cell'])
        ws[f'C{i}'].number_format = '0.0%'

        ws[f'D{i}'] = f'=IF(C{i}="","",IF(C{i}>=0.9,"Compliant",IF(C{i}>=0.7,"Partial","Non-Compliant")))'
        apply_cell_style(ws[f'D{i}'], styles['calculated_cell'])

    # --- Acknowledgment Rate KPI ---
    ws['A33'] = 'Policy Acknowledgment Rate:'
    apply_cell_style(ws['A33'], styles['label_cell'])
    apply_cell_style(ws['B33'], styles['input_cell'])
    ws['B33'].number_format = '0.0%'

    # Named range for cross-sheet reference
    ws['B24'] = '=B20'  # Link to overall score

    # Set column widths
    set_column_widths(ws, {'A': 35, 'B': 20, 'C': 15, 'D': 15, 'E': 15, 'F': 15})

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False


def create_roles_summary_sheet(wb, styles):
    """
    Sheet 5: Roles_Summary

    S2 (A.5.2) domain metrics from Roles & Responsibilities Assessment.
    """
    ws = wb.create_sheet("Roles_Summary")

    # --- Header ---
    merge_and_style(ws, 'A1:F1',
                   'ROLES & RESPONSIBILITIES SUMMARY (A.5.2)',
                   styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:F2',
                   'S2 Workbook Metrics - Information Security Roles and Responsibilities',
                   styles['header_sub'])

    # --- Source Reference ---
    ws['A4'] = 'Source Workbook:'
    apply_cell_style(ws['A4'], styles['label_cell'])
    ws['B4'] = 'ISMS-IMP-A.5.1-2-6.1-2.S2_Roles_Responsibilities'
    apply_cell_style(ws['B4'], styles['calculated_cell'])

    ws['A5'] = 'Weight in Dashboard:'
    apply_cell_style(ws['A5'], styles['label_cell'])
    ws['B5'] = '25%'
    apply_cell_style(ws['B5'], styles['calculated_cell'])

    # --- Key Metrics ---
    merge_and_style(ws, 'A7:F7', 'KEY METRICS', styles['section_header'])

    metrics = [
        (9, 'Total Roles Defined:', None, True),
        (10, 'Roles with Clear Ownership:', None, True),
        (11, 'Roles with RACI Matrix:', None, True),
        (12, 'Roles with Documented Responsibilities:', None, True),
        (13, 'Key Security Roles Filled:', None, True),
        (14, 'Roles with Succession Planning:', None, True),
        (15, 'Segregation of Duties Verified:', None, True),
        (16, 'Role-Based Access Aligned:', None, True),
        (17, 'Roles Requiring Update:', None, True),
        (18, 'Vacant Critical Roles:', None, True)
    ]

    for row_num, label, value, is_input in metrics:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        if is_input:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])
        else:
            ws[f'B{row_num}'] = value
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])

    # --- Overall Compliance Score ---
    merge_and_style(ws, 'A20:B20', 'OVERALL COMPLIANCE SCORE', styles['subsection_header'])
    ws['A20'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')

    apply_cell_style(ws['B20'], styles['input_cell'])
    ws['B20'].number_format = '0.0%'

    # --- Compliance Breakdown ---
    merge_and_style(ws, 'A22:F22', 'COMPLIANCE BY DOMAIN', styles['section_header'])

    domains = [
        ('Role Definition Completeness', '30%'),
        ('RACI Matrix Coverage', '25%'),
        ('Segregation of Duties', '20%'),
        ('Succession Planning', '15%'),
        ('Role Documentation', '10%')
    ]

    for col, header in zip(['A', 'B', 'C', 'D'], ['Domain', 'Weight', 'Score', 'Status']):
        ws[f'{col}24'] = header
        apply_cell_style(ws[f'{col}24'], styles['column_header'])

    for i, (domain, weight) in enumerate(domains, start=25):
        ws[f'A{i}'] = domain
        apply_cell_style(ws[f'A{i}'], styles['label_cell'])

        ws[f'B{i}'] = weight
        apply_cell_style(ws[f'B{i}'], styles['calculated_cell'])

        apply_cell_style(ws[f'C{i}'], styles['input_cell'])
        ws[f'C{i}'].number_format = '0.0%'

        ws[f'D{i}'] = f'=IF(C{i}="","",IF(C{i}>=0.9,"Compliant",IF(C{i}>=0.7,"Partial","Non-Compliant")))'
        apply_cell_style(ws[f'D{i}'], styles['calculated_cell'])

    # Set column widths
    set_column_widths(ws, {'A': 35, 'B': 20, 'C': 15, 'D': 15, 'E': 15, 'F': 15})

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False


def create_screening_summary_sheet(wb, styles):
    """
    Sheet 6: Screening_Summary

    S3 (A.6.1) domain metrics + FADP indicator.
    """
    ws = wb.create_sheet("Screening_Summary")

    # --- Header ---
    merge_and_style(ws, 'A1:F1',
                   'SCREENING & VETTING SUMMARY (A.6.1)',
                   styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:F2',
                   'S3 Workbook Metrics - Personnel Screening and Background Verification',
                   styles['header_sub'])

    # --- Source Reference ---
    ws['A4'] = 'Source Workbook:'
    apply_cell_style(ws['A4'], styles['label_cell'])
    ws['B4'] = 'ISMS-IMP-A.5.1-2-6.1-2.S3_Screening_Vetting'
    apply_cell_style(ws['B4'], styles['calculated_cell'])

    ws['A5'] = 'Weight in Dashboard:'
    apply_cell_style(ws['A5'], styles['label_cell'])
    ws['B5'] = '20%'
    apply_cell_style(ws['B5'], styles['calculated_cell'])

    # --- Key Metrics ---
    merge_and_style(ws, 'A7:F7', 'KEY METRICS', styles['section_header'])

    metrics = [
        (9, 'Total Personnel in Scope:', None, True),
        (10, 'Screening Completed:', None, True),
        (11, 'Screening Pending:', None, True),
        (12, 'Screening Overdue:', None, True),
        (13, 'Reference Checks Completed:', None, True),
        (14, 'Criminal Record Checks:', None, True),
        (15, 'Credential Verifications:', None, True),
        (16, 'Re-Screening Due:', None, True),
        (17, 'Exceptions Documented:', None, True),
        (18, 'Risk-Based Screening Applied:', None, True)
    ]

    for row_num, label, value, is_input in metrics:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        if is_input:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])
        else:
            ws[f'B{row_num}'] = value
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])

    # --- Overall Compliance Score ---
    merge_and_style(ws, 'A20:B20', 'OVERALL COMPLIANCE SCORE', styles['subsection_header'])

    apply_cell_style(ws['B20'], styles['input_cell'])
    ws['B20'].number_format = '0.0%'

    # --- FADP Compliance Section ---
    merge_and_style(ws, 'A22:F22', 'SWISS FADP COMPLIANCE', styles['section_header'])
    ws['A22'].fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')

    fadp_items = [
        (24, 'FADP Compliant Screening Process:', None),
        (25, 'Data Minimization Applied:', None),
        (26, 'FADP Training Completed:', None),
        (27, 'Consent Documentation:', None),
        (28, 'Cross-Border Transfer Controls:', None),
        (29, 'Data Retention Compliance:', None)
    ]

    for row_num, label, value in fadp_items:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # Data validation for FADP fields
    fadp_dv = create_data_validation(['Yes', 'Partial', 'No', 'N/A'])
    ws.add_data_validation(fadp_dv)
    fadp_dv.add('B24:B29')

    # --- Screening Completion Rate KPI ---
    ws['B24'] = '=B10/(B9)*100'  # This links to the screening rate
    ws['B24'].number_format = '0.0%'

    # Set column widths
    set_column_widths(ws, {'A': 35, 'B': 20, 'C': 15, 'D': 15, 'E': 15, 'F': 15})

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False


def create_contract_summary_sheet(wb, styles):
    """
    Sheet 7: Contract_Summary

    S4 (A.6.2/A.6.5) domain metrics + NDA focal point.
    """
    ws = wb.create_sheet("Contract_Summary")

    # --- Header ---
    merge_and_style(ws, 'A1:F1',
                   'EMPLOYMENT TERMS SUMMARY (A.6.2)',
                   styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:F2',
                   'S4 Workbook Metrics - Terms and Conditions of Employment',
                   styles['header_sub'])

    # --- Source Reference ---
    ws['A4'] = 'Source Workbook:'
    apply_cell_style(ws['A4'], styles['label_cell'])
    ws['B4'] = 'ISMS-IMP-A.5.1-2-6.1-2.S4_Employment_Terms'
    apply_cell_style(ws['B4'], styles['calculated_cell'])

    ws['A5'] = 'Weight in Dashboard:'
    apply_cell_style(ws['A5'], styles['label_cell'])
    ws['B5'] = '20%'
    apply_cell_style(ws['B5'], styles['calculated_cell'])

    # --- Key Metrics ---
    merge_and_style(ws, 'A7:F7', 'KEY METRICS', styles['section_header'])

    metrics = [
        (9, 'Total Employees in Scope:', None, True),
        (10, 'Contracts with Security Terms:', None, True),
        (11, 'Contracts Reviewed (Annual):', None, True),
        (12, 'Contracts Requiring Update:', None, True),
        (13, 'Security Clauses Standard:', None, True),
        (14, 'Acceptable Use Acknowledged:', None, True),
        (15, 'Post-Employment Terms Defined:', None, True),
        (16, 'Contractor Agreements Compliant:', None, True),
        (17, 'Third-Party Terms Verified:', None, True),
        (18, 'Contract Templates Updated:', None, True)
    ]

    for row_num, label, value, is_input in metrics:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        if is_input:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])
        else:
            ws[f'B{row_num}'] = value
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])

    # --- Overall Compliance Score ---
    merge_and_style(ws, 'A20:B20', 'OVERALL COMPLIANCE SCORE', styles['subsection_header'])

    apply_cell_style(ws['B20'], styles['input_cell'])
    ws['B20'].number_format = '0.0%'

    # --- NDA Compliance Section ---
    merge_and_style(ws, 'A22:F22', 'NDA COMPLIANCE (FOCAL POINT)', styles['section_header'])

    nda_items = [
        (24, 'NDA Compliance Rate:', None),
        (25, 'NDAs Signed (Employees):', None),
        (26, 'NDAs Signed (Contractors):', None),
        (27, 'NDAs Signed (Third Parties):', None),
        (28, 'NDAs Pending Signature:', None),
        (29, 'NDAs Expired/Requiring Renewal:', None),
        (30, 'NDA Template Last Updated:', None),
        (31, 'NDA Covers FADP Requirements:', None)
    ]

    for row_num, label, value in nda_items:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # Number format for rate
    ws['B24'].number_format = '0.0%'

    # Date format for template update
    ws['B30'].number_format = 'DD.MM.YYYY'

    # Data validation for FADP coverage
    fadp_dv = create_data_validation(['Yes', 'Partial', 'No'])
    ws.add_data_validation(fadp_dv)
    fadp_dv.add('B31')

    # --- Compliance Breakdown ---
    merge_and_style(ws, 'A33:F33', 'COMPLIANCE BY DOMAIN', styles['section_header'])

    domains = [
        ('Security Clause Coverage', '30%'),
        ('NDA Compliance', '25%'),
        ('Acceptable Use Acknowledgment', '20%'),
        ('Post-Employment Terms', '15%'),
        ('Contract Review Currency', '10%')
    ]

    for col, header in zip(['A', 'B', 'C', 'D'], ['Domain', 'Weight', 'Score', 'Status']):
        ws[f'{col}35'] = header
        apply_cell_style(ws[f'{col}35'], styles['column_header'])

    for i, (domain, weight) in enumerate(domains, start=36):
        ws[f'A{i}'] = domain
        apply_cell_style(ws[f'A{i}'], styles['label_cell'])

        ws[f'B{i}'] = weight
        apply_cell_style(ws[f'B{i}'], styles['calculated_cell'])

        apply_cell_style(ws[f'C{i}'], styles['input_cell'])
        ws[f'C{i}'].number_format = '0.0%'

        ws[f'D{i}'] = f'=IF(C{i}="","",IF(C{i}>=0.9,"Compliant",IF(C{i}>=0.7,"Partial","Non-Compliant")))'
        apply_cell_style(ws[f'D{i}'], styles['calculated_cell'])

    # Set column widths
    set_column_widths(ws, {'A': 35, 'B': 20, 'C': 15, 'D': 15, 'E': 15, 'F': 15})

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False


def create_gap_analysis_sheet(wb, styles):
    """
    Sheet 8: Gap_Analysis

    Consolidated gaps with cross-domain flags.
    """
    ws = wb.create_sheet("Gap_Analysis")

    # --- Header ---
    merge_and_style(ws, 'A1:N1', 'CONSOLIDATED GAP ANALYSIS', styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:N2',
                   'Cross-Domain Gap Identification and Remediation Tracking',
                   styles['header_sub'])

    # --- Summary Statistics ---
    merge_and_style(ws, 'A4:N4', 'GAP SUMMARY', styles['section_header'])

    summary_items = [
        (6, 'Total Gaps Identified:', '=COUNTA(A11:A110)-1'),
        (7, 'Critical Gaps:', '=COUNTIF(F11:F110,"Critical")'),
        (8, 'High Gaps:', '=COUNTIF(F11:F110,"High")'),
        (9, 'Medium Gaps:', '=COUNTIF(F11:F110,"Medium")'),
        (10, 'Low Gaps:', '=COUNTIF(F11:F110,"Low")')
    ]

    for row_num, label, formula in summary_items:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        ws[f'B{row_num}'] = formula
        apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])

    # Cross-domain gap count (for KPI reference)
    ws['A50'] = 'Cross-Domain Gap Count:'
    apply_cell_style(ws['A50'], styles['label_cell'])
    ws['B50'] = '=COUNTIF(H11:H110,"Yes")'
    apply_cell_style(ws['B50'], styles['calculated_cell'])

    ws['A51'] = 'Cross-Domain Gaps:'
    apply_cell_style(ws['A51'], styles['label_cell'])
    ws['B51'] = '=B50'
    apply_cell_style(ws['B51'], styles['calculated_cell'])

    # --- Column Headers ---
    headers = [
        ('A', 'Gap_ID', 12),
        ('B', 'Source_Workbook', 15),
        ('C', 'Primary_Control', 12),
        ('D', 'Gap_Title', 30),
        ('E', 'Gap_Description', 40),
        ('F', 'Risk_Level', 12),
        ('G', 'Affected_Domains', 20),
        ('H', 'Cross_Domain_Flag', 15),
        ('I', 'Remediation_Action', 35),
        ('J', 'Owner', 20),
        ('K', 'Target_Date', 12),
        ('L', 'Status', 15),
        ('M', 'FADP_Relevant', 12),
        ('N', 'Notes', 30)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}11'] = header_text
        apply_cell_style(ws[f'{col_letter}11'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (12-110, 100 gaps) ---
    for row in range(12, 111):
        # Gap_ID auto-generated
        ws[f'A{row}'] = f'=IF(D{row}<>"",TEXT(ROW()-11,"GAP-000"),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # All other columns manual entry
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'I', 'J', 'K', 'L', 'N']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # Cross-domain flag auto-calculated
        ws[f'H{row}'] = f'=IF(G{row}="","",IF(LEN(G{row})-LEN(SUBSTITUTE(G{row},",",""))>=1,"Yes","No"))'
        apply_cell_style(ws[f'H{row}'], styles['calculated_cell'])

        # FADP Relevant - input
        apply_cell_style(ws[f'M{row}'], styles['input_cell'])

    # Date formatting
    for row in range(12, 111):
        ws[f'K{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    source_dv = create_data_validation(['S1-Policy', 'S2-Roles', 'S3-Screening', 'S4-Contracts', 'Cross-Domain'])
    ws.add_data_validation(source_dv)
    source_dv.add('B12:B110')

    control_dv = create_data_validation(['A.5.1', 'A.5.2', 'A.6.1', 'A.6.2', 'Multiple'])
    ws.add_data_validation(control_dv)
    control_dv.add('C12:C110')

    risk_dv = create_data_validation(['Critical', 'High', 'Medium', 'Low'])
    ws.add_data_validation(risk_dv)
    risk_dv.add('F12:F110')

    status_dv = create_data_validation(['Open', 'In-Progress', 'Blocked', 'Closed', 'Accepted-Risk'])
    ws.add_data_validation(status_dv)
    status_dv.add('L12:L110')

    fadp_dv = create_data_validation(['Yes', 'No'])
    ws.add_data_validation(fadp_dv)
    fadp_dv.add('M12:M110')

    # Freeze panes
    ws.freeze_panes = 'A12'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_evidence_register_sheet(wb, styles):
    """
    Sheet 9: Evidence_Register

    Consolidated evidence index from all S1-S4 workbooks.
    """
    ws = wb.create_sheet("Evidence_Register")

    # --- Header ---
    merge_and_style(ws, 'A1:K1', 'CONSOLIDATED EVIDENCE REGISTER', styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:K2',
                   'Evidence Index for Governance Compliance Dashboard',
                   styles['header_sub'])

    # --- Column Headers ---
    headers = [
        ('A', 'Evidence_ID', 12),
        ('B', 'Source_Workbook', 15),
        ('C', 'Evidence_Type', 20),
        ('D', 'Description', 40),
        ('E', 'Related_Control', 12),
        ('F', 'Related_Gap_ID', 12),
        ('G', 'File_Location', 40),
        ('H', 'Date_Collected', 12),
        ('I', 'Collected_By', 20),
        ('J', 'Verification_Status', 15),
        ('K', 'Notes', 30)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (5-154, 150 evidence items) ---
    for row in range(5, 155):
        # Evidence_ID auto-generated
        ws[f'A{row}'] = f'=IF(D{row}<>"",TEXT(ROW()-4,"EVD-000"),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # All other columns manual entry
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

    # Date formatting
    for row in range(5, 155):
        ws[f'H{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    source_dv = create_data_validation(['S1-Policy', 'S2-Roles', 'S3-Screening', 'S4-Contracts', 'S5-Dashboard'])
    ws.add_data_validation(source_dv)
    source_dv.add('B5:B154')

    evidence_type_dv = create_data_validation([
        'Policy-Document', 'Approval-Record', 'Meeting-Minutes', 'Training-Record',
        'Screening-Report', 'Contract-Copy', 'NDA-Signed', 'Acknowledgment-Record',
        'Audit-Report', 'Assessment-Report', 'RACI-Matrix', 'Org-Chart', 'Other'
    ])
    ws.add_data_validation(evidence_type_dv)
    evidence_type_dv.add('C5:C154')

    control_dv = create_data_validation(['A.5.1', 'A.5.2', 'A.6.1', 'A.6.2', 'Multiple'])
    ws.add_data_validation(control_dv)
    control_dv.add('E5:E154')

    verification_dv = create_data_validation(['Verified', 'Pending', 'Not-Verified', 'Expired'])
    ws.add_data_validation(verification_dv)
    verification_dv.add('J5:J154')

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_action_items_sheet(wb, styles):
    """
    Sheet 10: Action_Items

    Remediation tracker linked to Gap_Analysis.
    """
    ws = wb.create_sheet("Action_Items")

    # --- Header ---
    merge_and_style(ws, 'A1:M1', 'ACTION ITEMS TRACKER', styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:M2',
                   'Remediation Tracking for Identified Gaps',
                   styles['header_sub'])

    # --- Summary ---
    merge_and_style(ws, 'A4:M4', 'ACTION SUMMARY', styles['section_header'])

    summary_items = [
        (6, 'Total Actions:', '=COUNTA(A10:A109)-1'),
        (7, 'Actions Open:', '=COUNTIF(G10:G109,"Open")+COUNTIF(G10:G109,"In-Progress")'),
        (8, 'Actions Blocked:', '=COUNTIF(G10:G109,"Blocked")'),
        (9, 'Actions Completed:', '=COUNTIF(G10:G109,"Completed")')
    ]

    for row_num, label, formula in summary_items:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        ws[f'B{row_num}'] = formula
        apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])

    # --- Column Headers ---
    headers = [
        ('A', 'Action_ID', 12),
        ('B', 'Related_Gap_ID', 12),
        ('C', 'Action_Description', 40),
        ('D', 'Priority', 12),
        ('E', 'Owner', 20),
        ('F', 'Target_Date', 12),
        ('G', 'Status', 15),
        ('H', 'Progress_%', 12),
        ('I', 'Last_Update', 12),
        ('J', 'Update_Notes', 30),
        ('K', 'Blocking_Issues', 30),
        ('L', 'Escalation_Required', 15),
        ('M', 'Escalation_To', 20)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}10'] = header_text
        apply_cell_style(ws[f'{col_letter}10'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (11-109, 100 actions) ---
    for row in range(11, 110):
        # Action_ID linked to Gap_ID
        ws[f'A{row}'] = f'=IF(B{row}<>"",SUBSTITUTE(B{row},"GAP","ACT"),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # All other columns manual entry
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

    # Formatting
    for row in range(11, 110):
        ws[f'F{row}'].number_format = 'DD.MM.YYYY'
        ws[f'H{row}'].number_format = '0%'
        ws[f'I{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    priority_dv = create_data_validation(['Critical', 'High', 'Medium', 'Low'])
    ws.add_data_validation(priority_dv)
    priority_dv.add('D11:D109')

    status_dv = create_data_validation(['Open', 'In-Progress', 'Blocked', 'Completed', 'Cancelled'])
    ws.add_data_validation(status_dv)
    status_dv.add('G11:G109')

    escalation_dv = create_data_validation(['Yes', 'No'])
    ws.add_data_validation(escalation_dv)
    escalation_dv.add('L11:L109')

    # Freeze panes
    ws.freeze_panes = 'A11'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_approval_signoff_sheet(wb, styles):
    """
    Sheet 11: Approval_Sign_Off

    Three-level governance certification with FADP acknowledgment.
    """
    ws = wb.create_sheet("Approval_Sign_Off")

    # --- Header ---
    merge_and_style(ws, 'A1:D1', 'GOVERNANCE CERTIFICATION', styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:D2',
                   'Three-Level Approval with FADP Acknowledgment',
                   styles['header_sub'])

    # --- Assessment Summary ---
    merge_and_style(ws, 'A4:D4', 'ASSESSMENT SUMMARY', styles['section_header'])

    summary_items = [
        (6, 'Document ID:', DOCUMENT_ID),
        (7, 'Assessment Period:', '=Executive_Summary!B6'),
        (8, 'Overall Compliance Score:', '=Executive_Summary!E6'),
        (9, 'Policy Domain (A.5.1):', '=Policy_Summary!B20'),
        (10, 'Roles Domain (A.5.2):', '=Roles_Summary!B20'),
        (11, 'Screening Domain (A.6.1):', '=Screening_Summary!B20'),
        (12, 'Employment Domain (A.6.2):', '=Contract_Summary!B20'),
        (13, 'Total Gaps Identified:', '=Gap_Analysis!B6'),
        (14, 'Critical/High Gaps:', '=Gap_Analysis!B7+Gap_Analysis!B8'),
        (15, 'Cross-Domain Gaps:', '=Gap_Analysis!B51'),
        (16, 'FADP Compliance Status:', '=Executive_Summary!I6')
    ]

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30

    for row_num, label, value in summary_items:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        ws[f'B{row_num}'] = value
        if value.startswith('='):
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])

    # Percentage formatting
    for row in [8, 9, 10, 11, 12]:
        ws[f'B{row}'].number_format = '0.0%'

    # --- Level 1: Prepared By ---
    merge_and_style(ws, 'A18:D18', 'LEVEL 1: PREPARED BY (Assessor)', styles['section_header'])

    level1_fields = [
        (20, 'Name:'),
        (21, 'Role:'),
        (22, 'Date:'),
        (23, 'Signature:'),
        (24, 'Certification:', 'I certify this assessment is complete and accurate based on S1-S4 workbook data')
    ]

    for row_num, label, *rest in level1_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])

        if rest:
            ws[f'B{row_num}'] = rest[0]
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    ws['B22'].number_format = 'DD.MM.YYYY'

    # --- Level 2: Reviewed By ---
    merge_and_style(ws, 'A27:D27', 'LEVEL 2: REVIEWED BY (Security Manager)', styles['section_header'])

    level2_fields = [
        (29, 'Name:'),
        (30, 'Role:'),
        (31, 'Date:'),
        (32, 'Signature:')
    ]

    for row_num, label in level2_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    ws['B31'].number_format = 'DD.MM.YYYY'

    # Review checklist
    ws['A34'] = 'Review Checklist:'
    apply_cell_style(ws['A34'], styles['label_cell'])

    checklist_items = [
        'S1-S4 workbook data verified',
        'Weighted scores calculated correctly',
        'All gaps accounted for',
        'Cross-domain impacts identified',
        'NDA compliance rate verified',
        'FADP indicators reviewed',
        'Evidence sufficiency confirmed'
    ]

    for i, item in enumerate(checklist_items, start=35):
        ws[f'A{i}'] = f'[ ] {item}'
        ws.merge_cells(f'A{i}:D{i}')

    ws['A42'] = 'Comments:'
    apply_cell_style(ws['A42'], styles['label_cell'])
    ws.merge_cells('B42:D43')
    apply_cell_style(ws['B42'], styles['input_cell'])

    # --- Level 3: Approved By (CISO) with FADP Acknowledgment ---
    merge_and_style(ws, 'A46:D46', 'LEVEL 3: APPROVED BY (CISO)', styles['section_header'])

    level3_fields = [
        (48, 'Name:'),
        (49, 'Role:'),
        (50, 'Date:'),
        (51, 'Signature:')
    ]

    for row_num, label in level3_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    ws['B50'].number_format = 'DD.MM.YYYY'

    # CISO Certification with FADP
    ws['A53'] = 'CISO Certification:'
    apply_cell_style(ws['A53'], styles['label_cell'])

    certification_text = (
        'I approve this Governance Compliance Dashboard as an accurate representation of '
        'the organization\'s compliance status for controls A.5.1, A.5.2, A.6.1, and A.6.2. '
        'I acknowledge the FADP compliance status and accept residual risk for identified gaps.'
    )
    ws.merge_cells('B53:D55')
    ws['B53'] = certification_text
    ws['B53'].alignment = Alignment(wrap_text=True, vertical='top')
    apply_cell_style(ws['B53'], styles['calculated_cell'])

    # FADP Acknowledgment
    merge_and_style(ws, 'A57:D57', 'FADP ACKNOWLEDGMENT', styles['section_header'])
    ws['A57'].fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')

    ws['A59'] = 'FADP Compliance Verified:'
    apply_cell_style(ws['A59'], styles['label_cell'])
    apply_cell_style(ws['B59'], styles['input_cell'])

    fadp_dv = create_data_validation(['Yes', 'No', 'Partial'])
    ws.add_data_validation(fadp_dv)
    fadp_dv.add('B59')

    ws['A60'] = 'Data Protection Officer Consulted:'
    apply_cell_style(ws['A60'], styles['label_cell'])
    apply_cell_style(ws['B60'], styles['input_cell'])

    consult_dv = create_data_validation(['Yes', 'No', 'N/A'])
    ws.add_data_validation(consult_dv)
    consult_dv.add('B60')

    ws['A61'] = 'FADP Risk Acceptance:'
    apply_cell_style(ws['A61'], styles['label_cell'])
    ws.merge_cells('B61:D62')
    apply_cell_style(ws['B61'], styles['input_cell'])

    # --- Assessment Status ---
    merge_and_style(ws, 'A64:D64', 'ASSESSMENT STATUS', styles['section_header'])

    ws['A66'] = 'Final Assessment Status:'
    apply_cell_style(ws['A66'], styles['label_cell'])
    apply_cell_style(ws['B66'], styles['input_cell'])

    status_dv = create_data_validation(['Draft', 'Under-Review', 'Approved', 'Audit-Ready'])
    ws.add_data_validation(status_dv)
    status_dv.add('B66')

    ws['A67'] = 'Next Review Date:'
    apply_cell_style(ws['A67'], styles['label_cell'])
    apply_cell_style(ws['B67'], styles['input_cell'])
    ws['B67'].number_format = 'DD.MM.YYYY'

    ws['A68'] = 'Board Presentation Date:'
    apply_cell_style(ws['A68'], styles['label_cell'])
    apply_cell_style(ws['B68'], styles['input_cell'])
    ws['B68'].number_format = 'DD.MM.YYYY'

    # Audit Readiness Checklist
    ws['A70'] = 'Audit Readiness:'
    apply_cell_style(ws['A70'], styles['label_cell'])

    audit_items = [
        'All three approval levels complete',
        'Evidence 100% verified',
        'All critical gaps have remediation plans',
        'FADP compliance acknowledged',
        'Board presentation scheduled'
    ]

    for i, item in enumerate(audit_items, start=71):
        ws[f'A{i}'] = f'[ ] {item}'
        ws.merge_cells(f'A{i}:D{i}')

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False


# ==============================================================================
# MAIN EXECUTION
# ==============================================================================

def main():
    """
    Main function to generate the Governance Compliance Dashboard workbook.

    Returns:
        int: 0 on success, 1 on failure
    """
    try:
        logger.info("=" * 80)
        logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME}")
        logger.info("Workbook Generator")
        logger.info("=" * 80)

        # Initialize workbook
        logger.info("Creating workbook...")
        wb = Workbook()
        styles = setup_styles()

        # Create all sheets
        logger.info("Creating Instructions_Legend...")
        create_instructions_legend_sheet(wb, styles)

        logger.info("Creating Executive_Summary...")
        create_executive_summary_sheet(wb, styles)

        logger.info("Creating Maturity_Assessment...")
        create_maturity_assessment_sheet(wb, styles)

        logger.info("Creating Policy_Summary...")
        create_policy_summary_sheet(wb, styles)

        logger.info("Creating Roles_Summary...")
        create_roles_summary_sheet(wb, styles)

        logger.info("Creating Screening_Summary...")
        create_screening_summary_sheet(wb, styles)

        logger.info("Creating Contract_Summary...")
        create_contract_summary_sheet(wb, styles)

        logger.info("Creating Gap_Analysis...")
        create_gap_analysis_sheet(wb, styles)

        logger.info("Creating Evidence_Register...")
        create_evidence_register_sheet(wb, styles)

        logger.info("Creating Action_Items...")
        create_action_items_sheet(wb, styles)

        logger.info("Creating Approval_Sign_Off...")
        create_approval_signoff_sheet(wb, styles)

        # Set workbook properties
        wb.properties.title = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
        wb.properties.subject = CONTROL_REF
        wb.properties.creator = "[Organization] Information Security Team"
        wb.properties.keywords = "Governance, Dashboard, ISMS, ISO27001, A.5.1, A.5.2, A.6.1, A.6.2, FADP"
        wb.properties.comments = f"Generated via Python script generate_a5_1_2_6_1_2_s5_governance_dashboard.py"

        # Save workbook
        logger.info(f"Saving workbook as: {OUTPUT_FILENAME}")
        wb.save(OUTPUT_FILENAME)

        logger.info("=" * 80)
        logger.info("Workbook generated successfully!")
        logger.info("=" * 80)
        logger.info("Sheet Summary:")
        logger.info("  1. Instructions_Legend - Navigation and color coding guide")
        logger.info("  2. Executive_Summary - Board-facing compliance overview")
        logger.info("  3. Maturity_Assessment - 5-level maturity model")
        logger.info("  4. Policy_Summary - S1 (A.5.1) metrics")
        logger.info("  5. Roles_Summary - S2 (A.5.2) metrics")
        logger.info("  6. Screening_Summary - S3 (A.6.1) metrics + FADP")
        logger.info("  7. Contract_Summary - S4 (A.6.2) metrics + NDA")
        logger.info("  8. Gap_Analysis - Consolidated gaps")
        logger.info("  9. Evidence_Register - Evidence index")
        logger.info(" 10. Action_Items - Remediation tracking")
        logger.info(" 11. Approval_Sign_Off - Three-level certification")
        logger.info("")
        logger.info("Dashboard Weights:")
        logger.info(f"  - Policy Framework (A.5.1): {int(WEIGHT_POLICY*100)}%")
        logger.info(f"  - Roles & Responsibilities (A.5.2): {int(WEIGHT_ROLES*100)}%")
        logger.info(f"  - Screening & Vetting (A.6.1): {int(WEIGHT_SCREENING*100)}%")
        logger.info(f"  - Employment Terms (A.6.2): {int(WEIGHT_EMPLOYMENT*100)}%")
        logger.info("")
        logger.info("Key Features:")
        logger.info("  - Aggregates S1-S4 findings")
        logger.info("  - Lifecycle-weighted scoring")
        logger.info("  - Cross-domain gap identification")
        logger.info("  - FADP compliance indicator")
        logger.info("  - NDA compliance rate KPI")
        logger.info("  - CISO certification with FADP acknowledgment")
        logger.info("")
        logger.info(f"File location: ./{OUTPUT_FILENAME}")

        return 0

    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
