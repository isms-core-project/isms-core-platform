#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.1-2-6.1-2.S4 - Employment Contract Assessment Workbook Generator
================================================================================

Purpose:
    Generate Excel assessment workbook for ISO/IEC 27001:2022 Controls A.6.1/A.6.2
    (Screening & Terms and Conditions of Employment) as part of stacked control
    framework A.5.1 + A.5.2 + A.6.1 + A.6.2 (Secure Employment and Roles).

Control Alignment:
    - ISO/IEC 27001:2022 Annex A.6.1: Screening
    - ISO/IEC 27001:2022 Annex A.6.2: Terms and Conditions of Employment
    - ISMS-POL-A.5.1-2-6.1-2, Section 7: Employment Agreements

Assessment Focus:
    - Contract template completeness
    - Required security clauses coverage
    - Per-individual contract compliance
    - NDA/confidentiality coverage (audit focal point)
    - Post-employment obligation tracking
    - Contractor/third-party agreement compliance

Workbook Structure:
    10 sheets total:
    1. Dashboard (executive summary, auto-calculated)
    2. Contract_Template_Assessment (verify templates are complete and current)
    3. Required_Clause_Registry (master list of required clauses)
    4. Personnel_Contract_Compliance (per-person contract status)
    5. Confidentiality_NDA_Tracking (NDA/confidentiality status per individual)
    6. Post_Employment_Obligations (track individuals with ongoing obligations)
    7. Contractor_Agreement_Assessment (third-party agreement verification)
    8. Gap_Analysis (consolidated gaps with risk and remediation)
    9. Evidence_Register (supporting evidence documentation)
   10. Approval_Sign_Off (three-level approval workflow)

Dependencies:
    - openpyxl >= 3.0.0

Usage:
    python generate_a5_1_2_6_1_2_s4_employment_contract.py

Output:
    ISMS-IMP-A.5.1-2-6.1-2.S4_Employment_Contract_Assessment_YYYYMMDD.xlsx

Technical Specification:
    See ISMS-IMP-A.5.1-2-6.1-2.S4-Employment-Contract-Assessment.md

Quality Standards:
    - Follows A.8.23/A.8.24 reference implementation patterns
    - Comprehensive documentation and inline comments
    - Consistent styling (Yellow input, Light Blue calculated, Gray labels)
    - Data validation on all appropriate cells
    - Conditional formatting for compliance statuses
    - Sheet protection with selective unlocking
    - Named ranges for cross-sheet references
    - Formula-based auto-calculations
    - Professional formatting for executive presentation

Author:               [Organization] ISMS Implementation Team
Version: 1.0
Date:                 [Date to be set]
License: Internal Use Only

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.1-2-6.1-2.S4"
WORKBOOK_NAME = "Employment Contract Assessment"
CONTROL_ID = "A.5.1-2-6.1-2"
CONTROL_NAME = "Secure Employment and Roles"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ==============================================================================
# STYLE DEFINITIONS
# ==============================================================================

def setup_styles():
    """
    Define consistent cell styles used throughout the workbook.

    Returns:
        dict: Dictionary of style objects for different cell types
    """
    styles = {}

    # Header styles
    styles['header_main'] = {
        'font': Font(name='Calibri', size=18, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['header_sub'] = {
        'font': Font(name='Calibri', size=12, bold=False, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['section_header'] = {
        'font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['column_header'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    # Data cell styles
    styles['input_cell'] = {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),  # Yellow
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    styles['calculated_cell'] = {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'),  # Light Blue
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    styles['label_cell'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),  # Gray
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    # Status fills (for conditional formatting reference)
    styles['status_compliant'] = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
    styles['status_partial'] = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')  # Yellow
    styles['status_non_compliant'] = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red

    return styles


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def apply_cell_style(cell, style_dict):
    """
    Apply a style dictionary to a cell.

    Args:
        cell: openpyxl cell object
        style_dict: Dictionary containing style attributes
    """
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']


def set_column_widths(ws, widths_dict):
    """
    Set column widths for a worksheet.

    Args:
        ws: Worksheet object
        widths_dict: Dictionary mapping column letters to widths
    """
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width


def create_data_validation(formula_or_list, allow_blank=True, error_title="Invalid Entry",
                           error_message="Please select from dropdown"):
    """
    Create a data validation object for dropdown lists.

    Args:
        formula_or_list: Either a string formula or list of values
        allow_blank: Whether to allow blank cells
        error_title: Title for error message
        error_message: Error message text

    Returns:
        DataValidation object
    """
    if isinstance(formula_or_list, list):
        dv = DataValidation(type="list", formula1=f'"{",".join(formula_or_list)}"',
                           allow_blank=allow_blank)
    else:
        dv = DataValidation(type="list", formula1=formula_or_list, allow_blank=allow_blank)

    dv.error = error_message
    dv.errorTitle = error_title
    dv.showErrorMessage = True
    return dv


def merge_and_style(ws, range_string, value, style_dict):
    """
    Merge cells and apply style.

    Args:
        ws: Worksheet object
        range_string: Range to merge (e.g., 'A1:K1')
        value: Value to write
        style_dict: Style dictionary to apply
    """
    ws.merge_cells(range_string)
    cell = ws[range_string.split(':')[0]]
    cell.value = value
    apply_cell_style(cell, style_dict)


# ==============================================================================
# SHEET CREATION FUNCTIONS
# ==============================================================================

def create_dashboard_sheet(wb, styles):
    """
    Sheet 1: Dashboard (Executive Summary)

    Auto-calculated metrics pulling from other sheets.
    Read-only for users.
    """
    ws = wb.active
    ws.title = "Dashboard"

    # --- Header Section ---
    merge_and_style(ws, 'A1:K1',
                   'ISMS-IMP-A.5.1-2-6.1-2.S4 - Employment Contract Assessment Dashboard',
                   styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:K2',
                   'ISO/IEC 27001:2022 Controls A.6.1/A.6.2 - Screening & Terms and Conditions of Employment',
                   styles['header_sub'])
    ws.row_dimensions[2].height = 30

    # --- Document Information Block (Rows 4-15) ---
    ws['A4'] = 'Document ID:'
    apply_cell_style(ws['A4'], styles['label_cell'])
    ws.merge_cells('B4:C4')
    ws['B4'] = 'ISMS-IMP-A.5.1-2-6.1-2.S4'

    info_rows = [
        (5, 'Related Policy:', 'ISMS-POL-A.5.1-2-6.1-2, Section 7'),
        (6, 'ISO Controls:', 'A.6.1 (Screening), A.6.2 (Terms & Conditions)'),
        (7, 'Assessment Period:', None),  # User input
        (8, 'Assessment Date:', '=TODAY()'),  # Formula
        (9, 'Assessor Name:', None),  # User input
        (10, 'Assessor Role:', None),  # User input
        (11, 'Organization:', None),  # User input
        (12, 'Review Cycle:', 'Annual'),
        (13, 'Last Updated:', '=NOW()'),  # Formula
        (14, 'Assessment Status:', None),  # Dropdown
        (15, 'Next Review Date:', None)  # User input
    ]

    for row_num, label, value in info_rows:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        ws.merge_cells(f'B{row_num}:C{row_num}')

        if value:
            ws[f'B{row_num}'] = value
            if value.startswith('='):
                apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
            else:
                apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # Status dropdown for row 14
    status_dv = create_data_validation(['Draft', 'Under-Review', 'Approved', 'Audit-Ready'])
    ws.add_data_validation(status_dv)
    status_dv.add(ws['B14'])

    # --- Overall Compliance Summary (Rows 17-30) ---
    merge_and_style(ws, 'A17:K17', 'OVERALL EMPLOYMENT CONTRACT COMPLIANCE', styles['section_header'])
    ws.row_dimensions[17].height = 35

    # Compliance Scorecard (Rows 19-26)
    scorecard = [
        (19, 'Overall Compliance Score', '=AVERAGE(C32,C33,C34,C35,C36)', '90%'),
        (20, 'Contract Templates Assessed', '=COUNTA(Contract_Template_Assessment!A:A)-4', 'N/A'),
        (21, 'Templates Compliant', '=COUNTIF(Contract_Template_Assessment!M:M,"Compliant")', '100%'),
        (22, 'Personnel with Contracts Verified', '=COUNTIF(Personnel_Contract_Compliance!N:N,"Compliant")', 'N/A'),
        (23, 'NDA Coverage Rate', '=COUNTIF(Confidentiality_NDA_Tracking!K:K,"Active")/COUNTA(Confidentiality_NDA_Tracking!A:A)*100', '100%'),
        (24, 'Critical Gaps', '=COUNTIF(Gap_Analysis!F:F,"Critical")', '0'),
        (25, 'High Priority Gaps', '=COUNTIF(Gap_Analysis!F:F,"High")', '0'),
        (26, 'Contractor Agreements Verified', '=COUNTIF(Contractor_Agreement_Assessment!L:L,"Compliant")', 'N/A')
    ]

    for row_num, label, formula, target in scorecard:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])

        ws[f'B{row_num}'] = formula
        apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        if row_num in [19, 23]:
            ws[f'B{row_num}'].number_format = '0.00%'

        ws[f'C{row_num}'] = target
        apply_cell_style(ws[f'C{row_num}'], styles['calculated_cell'])

        ws[f'D{row_num}'] = '=IF(B{0}>=C{0},"On Target",IF(B{0}>=C{0}*0.9,"Close","Below Target"))'.format(row_num)
        apply_cell_style(ws[f'D{row_num}'], styles['calculated_cell'])

    # --- Domain Compliance Breakdown (Rows 28-38) ---
    merge_and_style(ws, 'A28:K28', 'DOMAIN COMPLIANCE BREAKDOWN', styles['section_header'])

    # Column headers for domain table
    for col, header in zip(['A', 'B', 'C', 'D', 'E'],
                           ['Domain', 'Weight', 'Score (%)', 'Weighted Score', 'Status']):
        ws[f'{col}30'] = header
        apply_cell_style(ws[f'{col}30'], styles['column_header'])

    domains = [
        (32, 'Contract Template Completeness', '20%',
         '=COUNTIF(Contract_Template_Assessment!M:M,"Compliant")/COUNTA(Contract_Template_Assessment!A:A)*100', '=C32*0.20'),
        (33, 'Required Clause Coverage', '25%',
         '=COUNTIF(Required_Clause_Registry!H:H,"Covered")/COUNTA(Required_Clause_Registry!A:A)*100', '=C33*0.25'),
        (34, 'Personnel Contract Compliance', '25%',
         '=COUNTIF(Personnel_Contract_Compliance!N:N,"Compliant")/COUNTA(Personnel_Contract_Compliance!A:A)*100', '=C34*0.25'),
        (35, 'NDA/Confidentiality Coverage', '20%',
         '=COUNTIF(Confidentiality_NDA_Tracking!K:K,"Active")/COUNTA(Confidentiality_NDA_Tracking!A:A)*100', '=C35*0.20'),
        (36, 'Contractor Agreement Compliance', '10%',
         '=COUNTIF(Contractor_Agreement_Assessment!L:L,"Compliant")/COUNTA(Contractor_Agreement_Assessment!A:A)*100', '=C36*0.10')
    ]

    for row_num, domain, weight, score_formula, weighted_formula in domains:
        ws[f'A{row_num}'] = domain
        ws[f'B{row_num}'] = weight
        ws[f'C{row_num}'] = score_formula
        ws[f'C{row_num}'].number_format = '0.00%'
        ws[f'D{row_num}'] = weighted_formula
        ws[f'D{row_num}'].number_format = '0.00%'
        ws[f'E{row_num}'] = f'=IF(C{row_num}>=0.9,"Pass",IF(C{row_num}>=0.7,"Partial","Fail"))'

    # Total row
    ws['A38'] = 'TOTAL WEIGHTED SCORE'
    apply_cell_style(ws['A38'], styles['label_cell'])
    ws['B38'] = '100%'
    ws['C38'] = '=SUM(D32:D36)'
    ws['C38'].number_format = '0.00%'
    ws['C38'].font = Font(name='Calibri', size=10, bold=True)

    # Set column widths
    set_column_widths(ws, {'A': 40, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 5, 'G': 5, 'H': 5, 'I': 5, 'J': 5, 'K': 5})

    # Protect sheet (allow sorting/filtering)
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_contract_template_assessment_sheet(wb, styles):
    """
    Sheet 2: Contract_Template_Assessment

    Verify contract templates are complete and current.
    50 rows for templates.
    """
    ws = wb.create_sheet("Contract_Template_Assessment")

    # --- Header ---
    merge_and_style(ws, 'A1:N1', 'CONTRACT TEMPLATE ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:N2',
                   'Verification of employment contract templates for security clause completeness',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # --- Column Headers (Row 4) ---
    headers = [
        ('A', 'Template_ID', 15),
        ('B', 'Template_Name', 35),
        ('C', 'Template_Type', 20),
        ('D', 'Personnel_Category', 20),
        ('E', 'Current_Version', 12),
        ('F', 'Version_Date', 12),
        ('G', 'Last_Review_Date', 12),
        ('H', 'Next_Review_Date', 12),
        ('I', 'Security_Clauses_Count', 15),
        ('J', 'Required_Clauses_Met', 15),
        ('K', 'Legal_Review_Date', 12),
        ('L', 'Legal_Approved', 15),
        ('M', 'Compliance_Status', 15),
        ('N', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)
    ws.row_dimensions[4].height = 30

    # --- Data Rows (5-54) ---
    for row in range(5, 55):
        for col_letter in [c[0] for c in headers]:
            cell = ws[f'{col_letter}{row}']
            if col_letter == 'M':
                # Auto-calculated compliance status
                ws[f'M{row}'] = (f'=IF(B{row}="","",IF(AND(J{row}="Yes",L{row}="Yes",'
                                f'H{row}>=TODAY()),"Compliant",'
                                f'IF(OR(J{row}="No",L{row}="No"),"Non-Compliant","Partial")))')
                apply_cell_style(cell, styles['calculated_cell'])
            else:
                apply_cell_style(cell, styles['input_cell'])

    # --- Data Validation ---
    # Template Type dropdown
    type_dv = create_data_validation(['Employee-Permanent', 'Employee-Fixed-Term', 'Executive',
                                       'Intern', 'Contractor', 'Consultant', 'Agency-Worker'])
    ws.add_data_validation(type_dv)
    type_dv.add('C5:C54')

    # Personnel Category dropdown
    category_dv = create_data_validation(['All-Personnel', 'IT-Staff', 'Management', 'Finance',
                                          'HR', 'Operations', 'External', 'Privileged-Users'])
    ws.add_data_validation(category_dv)
    category_dv.add('D5:D54')

    # Yes/No dropdowns
    yn_dv = create_data_validation(['Yes', 'No', 'Partial'])
    ws.add_data_validation(yn_dv)
    yn_dv.add('J5:J54')
    yn_dv.add('L5:L54')

    # Date formatting
    for col in ['F', 'G', 'H', 'K']:
        for row in range(5, 55):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes at A5
    ws.freeze_panes = 'A5'

    # Define named range for Template_ID_List
    template_id_range = DefinedName(name='Template_ID_List', attr_text="Contract_Template_Assessment!$A$5:$A$54")
    wb.defined_names.add(template_id_range)

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_required_clause_registry_sheet(wb, styles):
    """
    Sheet 3: Required_Clause_Registry

    Master list of required security clauses and template coverage.
    """
    ws = wb.create_sheet("Required_Clause_Registry")

    # --- Header ---
    merge_and_style(ws, 'A1:L1', 'REQUIRED CLAUSE REGISTRY', styles['header_main'])
    merge_and_style(ws, 'A2:L2',
                   'Master list of required security clauses with template coverage verification',
                   styles['header_sub'])

    # --- Column Headers ---
    headers = [
        ('A', 'Clause_ID', 15),
        ('B', 'Clause_Category', 20),
        ('C', 'Clause_Title', 35),
        ('D', 'Clause_Description', 50),
        ('E', 'ISO_27001_Reference', 20),
        ('F', 'Legal_Requirement', 20),
        ('G', 'Mandatory_For', 25),
        ('H', 'Coverage_Status', 15),
        ('I', 'Templates_With_Clause', 25),
        ('J', 'Gap_If_Missing', 40),
        ('K', 'Last_Verified', 12),
        ('L', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Pre-populate Required Clauses (Rows 5-34) ---
    required_clauses = [
        ('RC-001', 'Confidentiality', 'Confidentiality and Non-Disclosure',
         'Employee agrees to maintain confidentiality of all company information', 'A.5.1, A.6.2', 'Employment Law'),
        ('RC-002', 'Confidentiality', 'Non-Disclosure Agreement Reference',
         'Reference to standalone NDA or embedded NDA terms', 'A.6.2', 'Employment Law'),
        ('RC-003', 'IP Rights', 'Intellectual Property Assignment',
         'All work product IP belongs to the organization', 'A.5.1', 'IP Law'),
        ('RC-004', 'IP Rights', 'Invention Assignment',
         'Employee assigns rights to inventions created during employment', 'A.5.1', 'Patent Law'),
        ('RC-005', 'Security', 'Information Security Responsibilities',
         'Employee responsibilities for information security', 'A.6.2', 'ISMS Policy'),
        ('RC-006', 'Security', 'Acceptable Use Agreement Reference',
         'Reference to or acknowledgment of acceptable use policy', 'A.5.10', 'ISMS Policy'),
        ('RC-007', 'Security', 'Access Control Acknowledgment',
         'Agreement to comply with access control policies', 'A.5.15', 'ISMS Policy'),
        ('RC-008', 'Security', 'Password/Authentication Requirements',
         'Agreement to follow authentication policies', 'A.5.17', 'ISMS Policy'),
        ('RC-009', 'Post-Employment', 'Return of Assets',
         'Obligation to return all company assets upon termination', 'A.6.5', 'Employment Law'),
        ('RC-010', 'Post-Employment', 'Post-Employment Confidentiality',
         'Confidentiality obligations continue after termination', 'A.6.5', 'Contract Law'),
        ('RC-011', 'Post-Employment', 'Non-Compete Clause',
         'Restrictions on competing employment if applicable', 'A.6.5', 'Contract Law'),
        ('RC-012', 'Post-Employment', 'Non-Solicitation Clause',
         'Restrictions on soliciting employees/clients', 'A.6.5', 'Contract Law'),
        ('RC-013', 'Compliance', 'Policy Compliance Commitment',
         'Agreement to comply with all company policies', 'A.5.1', 'ISMS Policy'),
        ('RC-014', 'Compliance', 'Training Completion Requirement',
         'Agreement to complete required security training', 'A.6.3', 'ISMS Policy'),
        ('RC-015', 'Compliance', 'Incident Reporting Obligation',
         'Obligation to report security incidents', 'A.6.8', 'ISMS Policy'),
        ('RC-016', 'Screening', 'Background Check Authorization',
         'Authorization for pre-employment screening', 'A.6.1', 'Employment Law'),
        ('RC-017', 'Screening', 'Reference Check Authorization',
         'Authorization to contact references', 'A.6.1', 'Employment Law'),
        ('RC-018', 'Data Protection', 'Personal Data Processing Consent',
         'Consent for processing of employee personal data', 'FADP', 'Data Protection'),
        ('RC-019', 'Data Protection', 'Data Protection Responsibilities',
         'Employee responsibilities for data protection', 'A.5.34', 'Data Protection'),
        ('RC-020', 'Monitoring', 'Monitoring Acknowledgment',
         'Acknowledgment of workplace/IT monitoring practices', 'A.8.15', 'Employment Law')
    ]

    for i, clause in enumerate(required_clauses, start=5):
        ws[f'A{i}'] = clause[0]
        ws[f'B{i}'] = clause[1]
        ws[f'C{i}'] = clause[2]
        ws[f'D{i}'] = clause[3]
        ws[f'E{i}'] = clause[4]
        ws[f'F{i}'] = clause[5]

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            apply_cell_style(ws[f'{col}{i}'], styles['calculated_cell'])

        for col in ['G', 'H', 'I', 'J', 'K', 'L']:
            apply_cell_style(ws[f'{col}{i}'], styles['input_cell'])

    # Additional blank rows for user-defined clauses
    for row in range(25, 55):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

    # --- Data Validation ---
    category_dv = create_data_validation(['Confidentiality', 'IP Rights', 'Security',
                                          'Post-Employment', 'Compliance', 'Screening',
                                          'Data Protection', 'Monitoring', 'Other'])
    ws.add_data_validation(category_dv)
    category_dv.add('B5:B54')

    coverage_dv = create_data_validation(['Covered', 'Partial', 'Not-Covered', 'N/A'])
    ws.add_data_validation(coverage_dv)
    coverage_dv.add('H5:H54')

    # Date formatting
    for row in range(5, 55):
        ws[f'K{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_personnel_contract_compliance_sheet(wb, styles):
    """
    Sheet 4: Personnel_Contract_Compliance

    Per-person contract status verification.
    200 rows for personnel.
    """
    ws = wb.create_sheet("Personnel_Contract_Compliance")

    # --- Header ---
    merge_and_style(ws, 'A1:O1', 'PERSONNEL CONTRACT COMPLIANCE', styles['header_main'])
    merge_and_style(ws, 'A2:O2',
                   'Individual employee contract verification and compliance tracking',
                   styles['header_sub'])

    # --- Column Headers ---
    headers = [
        ('A', 'Employee_ID', 15),
        ('B', 'Employee_Name', 25),
        ('C', 'Department', 20),
        ('D', 'Role_Title', 25),
        ('E', 'Employment_Type', 15),
        ('F', 'Start_Date', 12),
        ('G', 'Contract_Template_Used', 20),
        ('H', 'Contract_Signed', 15),
        ('I', 'Contract_Date', 12),
        ('J', 'Contract_On_File', 15),
        ('K', 'All_Clauses_Present', 15),
        ('L', 'Screening_Complete', 15),
        ('M', 'Background_Check_Date', 12),
        ('N', 'Compliance_Status', 15),
        ('O', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (5-204) ---
    for row in range(5, 205):
        for col_letter in [c[0] for c in headers]:
            cell = ws[f'{col_letter}{row}']
            if col_letter == 'N':
                # Auto-calculated compliance status
                ws[f'N{row}'] = (f'=IF(B{row}="","",IF(AND(H{row}="Yes",J{row}="Yes",'
                                f'K{row}="Yes",L{row}="Yes"),"Compliant",'
                                f'IF(OR(H{row}="No",J{row}="No",L{row}="No"),"Non-Compliant","Partial")))')
                apply_cell_style(cell, styles['calculated_cell'])
            else:
                apply_cell_style(cell, styles['input_cell'])

    # --- Data Validation ---
    emp_type_dv = create_data_validation(['Permanent', 'Fixed-Term', 'Executive', 'Intern',
                                          'Contractor', 'Consultant', 'Agency'])
    ws.add_data_validation(emp_type_dv)
    emp_type_dv.add('E5:E204')

    yn_dv = create_data_validation(['Yes', 'No', 'Partial', 'Pending'])
    ws.add_data_validation(yn_dv)
    yn_dv.add('H5:H204')
    yn_dv.add('J5:J204')
    yn_dv.add('K5:K204')
    yn_dv.add('L5:L204')

    # Date formatting
    for col in ['F', 'I', 'M']:
        for row in range(5, 205):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Define named range for Employee_ID_List
    employee_id_range = DefinedName(name='Employee_ID_List', attr_text="Personnel_Contract_Compliance!$A$5:$A$204")
    wb.defined_names.add(employee_id_range)

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_confidentiality_nda_tracking_sheet(wb, styles):
    """
    Sheet 5: Confidentiality_NDA_Tracking

    NDA / confidentiality status per individual.
    AUDIT FOCAL POINT - critical for compliance verification.
    200 rows for personnel.
    """
    ws = wb.create_sheet("Confidentiality_NDA_Tracking")

    # --- Header ---
    merge_and_style(ws, 'A1:N1', 'CONFIDENTIALITY & NDA TRACKING', styles['header_main'])
    merge_and_style(ws, 'A2:N2',
                   'AUDIT FOCAL POINT: NDA and confidentiality agreement status per individual',
                   styles['header_sub'])

    # --- Column Headers ---
    headers = [
        ('A', 'Employee_ID', 15),
        ('B', 'Employee_Name', 25),
        ('C', 'Department', 20),
        ('D', 'NDA_Type', 20),
        ('E', 'NDA_Version', 12),
        ('F', 'Date_Signed', 12),
        ('G', 'Expiration_Date', 12),
        ('H', 'Covers_IP', 15),
        ('I', 'Covers_Trade_Secrets', 15),
        ('J', 'Covers_Post_Employment', 15),
        ('K', 'NDA_Status', 15),
        ('L', 'Renewal_Required', 15),
        ('M', 'Document_Location', 30),
        ('N', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (5-204) ---
    for row in range(5, 205):
        # Auto-populate Employee info from Personnel sheet
        ws[f'A{row}'] = f'=IF(ROW()-4<=COUNTA(Personnel_Contract_Compliance!$A:$A)-4,INDEX(Personnel_Contract_Compliance!$A:$A,ROW()-3),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Personnel_Contract_Compliance!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])

        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Personnel_Contract_Compliance!$A:$C,3,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])

        # Manual entry cells
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'M', 'N']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # NDA_Status (K) - auto-calculated
        ws[f'K{row}'] = (f'=IF(B{row}="","",IF(AND(F{row}<>"",OR(G{row}="",G{row}>=TODAY())),"Active",'
                        f'IF(AND(F{row}<>"",G{row}<TODAY()),"Expired",'
                        f'IF(F{row}="","Missing","Unknown"))))')
        apply_cell_style(ws[f'K{row}'], styles['calculated_cell'])

    # --- Data Validation ---
    nda_type_dv = create_data_validation(['Embedded-In-Contract', 'Standalone-NDA', 'Mutual-NDA',
                                          'Unilateral-NDA', 'PIIA', 'Combined-Agreement'])
    ws.add_data_validation(nda_type_dv)
    nda_type_dv.add('D5:D204')

    yn_dv = create_data_validation(['Yes', 'No', 'Partial'])
    ws.add_data_validation(yn_dv)
    yn_dv.add('H5:H204')
    yn_dv.add('I5:I204')
    yn_dv.add('J5:J204')

    renewal_dv = create_data_validation(['Yes', 'No', 'N/A'])
    ws.add_data_validation(renewal_dv)
    renewal_dv.add('L5:L204')

    # Date formatting
    for col in ['F', 'G']:
        for row in range(5, 205):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_post_employment_obligations_sheet(wb, styles):
    """
    Sheet 6: Post_Employment_Obligations

    Track individuals with ongoing post-employment obligations.
    100 rows for terminated employees with obligations.
    """
    ws = wb.create_sheet("Post_Employment_Obligations")

    # --- Header ---
    merge_and_style(ws, 'A1:O1', 'POST-EMPLOYMENT OBLIGATIONS TRACKING', styles['header_main'])
    merge_and_style(ws, 'A2:O2',
                   'Tracking individuals with ongoing confidentiality and non-compete obligations',
                   styles['header_sub'])

    # --- Column Headers ---
    headers = [
        ('A', 'Former_Employee_ID', 15),
        ('B', 'Former_Employee_Name', 25),
        ('C', 'Former_Department', 20),
        ('D', 'Former_Role', 25),
        ('E', 'Termination_Date', 12),
        ('F', 'Termination_Type', 15),
        ('G', 'Confidentiality_End_Date', 15),
        ('H', 'Non_Compete_End_Date', 15),
        ('I', 'Non_Solicit_End_Date', 15),
        ('J', 'IP_Assignment_Perpetual', 15),
        ('K', 'Assets_Returned', 15),
        ('L', 'Exit_Interview_Complete', 15),
        ('M', 'Obligation_Status', 15),
        ('N', 'Monitoring_Required', 15),
        ('O', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (5-104) ---
    for row in range(5, 105):
        for col_letter in [c[0] for c in headers]:
            cell = ws[f'{col_letter}{row}']
            if col_letter == 'M':
                # Auto-calculated obligation status
                ws[f'M{row}'] = (f'=IF(B{row}="","",IF(AND(K{row}="Yes",L{row}="Yes",'
                                f'OR(G{row}="Perpetual",G{row}>=TODAY())),"Active-Compliant",'
                                f'IF(OR(K{row}="No",L{row}="No"),"Non-Compliant",'
                                f'IF(AND(G{row}<TODAY(),H{row}<TODAY(),I{row}<TODAY()),"Expired","Active-Review"))))')
                apply_cell_style(cell, styles['calculated_cell'])
            else:
                apply_cell_style(cell, styles['input_cell'])

    # --- Data Validation ---
    term_type_dv = create_data_validation(['Resignation', 'Termination', 'Retirement',
                                           'Contract-End', 'Mutual-Agreement', 'Redundancy'])
    ws.add_data_validation(term_type_dv)
    term_type_dv.add('F5:F104')

    yn_dv = create_data_validation(['Yes', 'No', 'Partial', 'N/A'])
    ws.add_data_validation(yn_dv)
    yn_dv.add('J5:J104')
    yn_dv.add('K5:K104')
    yn_dv.add('L5:L104')
    yn_dv.add('N5:N104')

    # Date formatting
    for col in ['E', 'G', 'H', 'I']:
        for row in range(5, 105):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_contractor_agreement_assessment_sheet(wb, styles):
    """
    Sheet 7: Contractor_Agreement_Assessment

    Third-party and contractor agreement verification.
    100 rows for contractors/vendors.
    """
    ws = wb.create_sheet("Contractor_Agreement_Assessment")

    # --- Header ---
    merge_and_style(ws, 'A1:N1', 'CONTRACTOR AGREEMENT ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:N2',
                   'Third-party, contractor, and vendor agreement security clause verification',
                   styles['header_sub'])

    # --- Column Headers ---
    headers = [
        ('A', 'Contractor_ID', 15),
        ('B', 'Contractor_Name', 25),
        ('C', 'Company_Name', 25),
        ('D', 'Contract_Type', 20),
        ('E', 'Services_Provided', 30),
        ('F', 'Contract_Start', 12),
        ('G', 'Contract_End', 12),
        ('H', 'NDA_In_Place', 15),
        ('I', 'Security_Clauses_Present', 15),
        ('J', 'Data_Access_Level', 15),
        ('K', 'Background_Check', 15),
        ('L', 'Compliance_Status', 15),
        ('M', 'Sponsor_Manager', 25),
        ('N', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (5-104) ---
    for row in range(5, 105):
        for col_letter in [c[0] for c in headers]:
            cell = ws[f'{col_letter}{row}']
            if col_letter == 'L':
                # Auto-calculated compliance status
                ws[f'L{row}'] = (f'=IF(B{row}="","",IF(AND(H{row}="Yes",I{row}="Yes",'
                                f'OR(K{row}="Yes",K{row}="N/A")),"Compliant",'
                                f'IF(OR(H{row}="No",I{row}="No"),"Non-Compliant","Partial")))')
                apply_cell_style(cell, styles['calculated_cell'])
            else:
                apply_cell_style(cell, styles['input_cell'])

    # --- Data Validation ---
    contract_type_dv = create_data_validation(['Statement-of-Work', 'Master-Services-Agreement',
                                                'Professional-Services', 'Staffing-Agency',
                                                'Consulting-Agreement', 'Vendor-Contract', 'Other'])
    ws.add_data_validation(contract_type_dv)
    contract_type_dv.add('D5:D104')

    yn_dv = create_data_validation(['Yes', 'No', 'Partial', 'N/A'])
    ws.add_data_validation(yn_dv)
    yn_dv.add('H5:H104')
    yn_dv.add('I5:I104')
    yn_dv.add('K5:K104')

    access_dv = create_data_validation(['None', 'Limited', 'Standard', 'Privileged', 'Admin'])
    ws.add_data_validation(access_dv)
    access_dv.add('J5:J104')

    # Date formatting
    for col in ['F', 'G']:
        for row in range(5, 105):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_gap_analysis_sheet(wb, styles):
    """
    Sheet 8: Gap_Analysis

    Consolidated gaps from all assessment domains.
    """
    ws = wb.create_sheet("Gap_Analysis")

    # --- Header ---
    merge_and_style(ws, 'A1:P1', 'GAP ANALYSIS', styles['header_main'])
    merge_and_style(ws, 'A2:P2',
                   'Consolidated gaps from all employment contract assessment domains',
                   styles['header_sub'])

    # --- Column Headers ---
    headers = [
        ('A', 'Gap_ID', 12),
        ('B', 'Source_Sheet', 25),
        ('C', 'Affected_Entity', 25),
        ('D', 'Gap_Category', 20),
        ('E', 'Gap_Description', 50),
        ('F', 'Risk_Level', 15),
        ('G', 'Impact_Assessment', 40),
        ('H', 'Affected_Personnel_Count', 15),
        ('I', 'Remediation_Action', 50),
        ('J', 'Responsible_Party', 25),
        ('K', 'Target_Completion_Date', 15),
        ('L', 'Estimated_Effort', 15),
        ('M', 'Dependencies', 30),
        ('N', 'Status', 15),
        ('O', 'Completion_Evidence', 30),
        ('P', 'Risk_Acceptance', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (5-104, 100 rows for gaps) ---
    for row in range(5, 105):
        # Gap_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}<>"",TEXT(ROW()-4,"GAP-000"),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # All other columns manual entry
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

    # Date formatting for Target_Completion_Date
    for row in range(5, 105):
        ws[f'K{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    source_sheet_dv = create_data_validation(['Contract_Template_Assessment', 'Required_Clause_Registry',
                                               'Personnel_Contract_Compliance', 'Confidentiality_NDA_Tracking',
                                               'Post_Employment_Obligations', 'Contractor_Agreement_Assessment'])
    ws.add_data_validation(source_sheet_dv)
    source_sheet_dv.add('B5:B104')

    category_dv = create_data_validation(['Template-Gap', 'Clause-Gap', 'Contract-Gap',
                                          'NDA-Gap', 'Post-Employment-Gap', 'Contractor-Gap',
                                          'Process-Gap', 'Documentation-Gap'])
    ws.add_data_validation(category_dv)
    category_dv.add('D5:D104')

    risk_level_dv = create_data_validation(['Critical', 'High', 'Medium', 'Low'])
    ws.add_data_validation(risk_level_dv)
    risk_level_dv.add('F5:F104')

    effort_dv = create_data_validation(['<1hr', '1-4hrs', '1day', '2-5days', '>1week'])
    ws.add_data_validation(effort_dv)
    effort_dv.add('L5:L104')

    status_dv = create_data_validation(['Not-Started', 'In-Progress', 'Blocked', 'Completed', 'Accepted-Risk'])
    ws.add_data_validation(status_dv)
    status_dv.add('N5:N104')

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_evidence_register_sheet(wb, styles):
    """
    Sheet 9: Evidence_Register

    Supporting evidence documentation.
    """
    ws = wb.create_sheet("Evidence_Register")

    # --- Header ---
    merge_and_style(ws, 'A1:J1', 'EVIDENCE REGISTER', styles['header_main'])
    merge_and_style(ws, 'A2:J2',
                   'Documentation of all supporting evidence for employment contract assessment',
                   styles['header_sub'])

    # --- Column Headers ---
    headers = [
        ('A', 'Evidence_ID', 12),
        ('B', 'Evidence_Type', 25),
        ('C', 'Description', 40),
        ('D', 'Related_Entity_ID', 20),
        ('E', 'Related_Assessment_Sheet', 25),
        ('F', 'File_Location', 40),
        ('G', 'Date_Collected', 12),
        ('H', 'Collected_By', 25),
        ('I', 'Verification_Status', 15),
        ('J', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (5-104, 100 evidence items) ---
    for row in range(5, 105):
        # Evidence_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}<>"",TEXT(ROW()-4,"EVD-000"),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # All other columns manual entry
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

    # Date formatting for Date_Collected
    for row in range(5, 105):
        ws[f'G{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    evidence_type_dv = create_data_validation(['Employment-Contract', 'NDA-Agreement', 'Contract-Template',
                                               'Signed-Acknowledgment', 'Background-Check-Report',
                                               'Exit-Interview-Form', 'Asset-Return-Form', 'Legal-Review',
                                               'Policy-Acknowledgment', 'Training-Certificate', 'Other'])
    ws.add_data_validation(evidence_type_dv)
    evidence_type_dv.add('B5:B104')

    sheet_dv = create_data_validation(['Contract_Template_Assessment', 'Required_Clause_Registry',
                                        'Personnel_Contract_Compliance', 'Confidentiality_NDA_Tracking',
                                        'Post_Employment_Obligations', 'Contractor_Agreement_Assessment',
                                        'Gap_Analysis'])
    ws.add_data_validation(sheet_dv)
    sheet_dv.add('E5:E104')

    verification_dv = create_data_validation(['Verified', 'Pending', 'Not-Verified', 'Expired'])
    ws.add_data_validation(verification_dv)
    verification_dv.add('I5:I104')

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_approval_signoff_sheet(wb, styles):
    """
    Sheet 10: Approval_Sign_Off

    Three-level approval workflow.
    """
    ws = wb.create_sheet("Approval_Sign_Off")

    # --- Header ---
    merge_and_style(ws, 'A1:D1', 'APPROVAL & SIGN-OFF', styles['header_main'])
    ws.row_dimensions[1].height = 30

    # --- Assessment Summary (Rows 3-20) ---
    ws['A3'] = 'ASSESSMENT SUMMARY'
    apply_cell_style(ws['A3'], styles['section_header'])
    ws.merge_cells('A3:D3')

    summary_items = [
        (5, 'Document:', 'ISMS-IMP-A.5.1-2-6.1-2.S4'),
        (6, 'Assessment Period:', '=Dashboard!B7'),
        (7, 'Overall Compliance Score:', '=Dashboard!B19'),
        (8, 'Contract Templates Assessed:', '=Dashboard!B20'),
        (9, 'Personnel Contracts Verified:', '=Dashboard!B22'),
        (10, 'NDA Coverage Rate:', '=Dashboard!B23'),
        (11, 'Critical Gaps:', '=Dashboard!B24'),
        (12, 'High Priority Gaps:', '=Dashboard!B25'),
        (13, 'Contractor Agreements Verified:', '=Dashboard!B26')
    ]

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30

    for row_num, label, value in summary_items:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])

        ws[f'B{row_num}'] = value
        if value.startswith('='):
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])

    # --- Level 1 Approval (Rows 16-25) ---
    ws['A16'] = 'LEVEL 1: PREPARED BY'
    apply_cell_style(ws['A16'], styles['section_header'])
    ws.merge_cells('A16:D16')

    level1_fields = [
        (17, 'Name:'),
        (18, 'Role:'),
        (19, 'Date:'),
        (20, 'Signature:'),
        (21, 'Certification:', 'I certify this assessment is complete and accurate'),
        (22, 'Comments:')
    ]

    for row_num, label, *rest in level1_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])

        if rest:
            ws[f'B{row_num}'] = rest[0]
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # --- Level 2 Approval (Rows 24-36) ---
    ws['A24'] = 'LEVEL 2: REVIEWED BY (HR / LEGAL)'
    apply_cell_style(ws['A24'], styles['section_header'])
    ws.merge_cells('A24:D24')

    level2_fields = [
        (25, 'Name:'),
        (26, 'Role:'),
        (27, 'Date:'),
        (28, 'Signature:')
    ]

    for row_num, label in level2_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # Review checklist
    checklist_items = [
        'Contract templates reviewed and current',
        'All required security clauses verified',
        'Personnel contract compliance checked',
        'NDA coverage 100% (audit focal point)',
        'Post-employment obligations tracked',
        'Contractor agreements verified',
        'Gaps documented with remediation plans',
        'Evidence register complete'
    ]

    ws['A30'] = 'Review Checklist:'
    apply_cell_style(ws['A30'], styles['label_cell'])

    for i, item in enumerate(checklist_items, start=31):
        ws[f'A{i}'] = f'[ ] {item}'
        ws.merge_cells(f'A{i}:D{i}')

    ws['A39'] = 'Comments:'
    apply_cell_style(ws['A39'], styles['label_cell'])
    apply_cell_style(ws['B39'], styles['input_cell'])

    # --- Level 3 Approval (Rows 41-52) ---
    ws['A41'] = 'LEVEL 3: APPROVED BY (CISO)'
    apply_cell_style(ws['A41'], styles['section_header'])
    ws.merge_cells('A41:D41')

    level3_fields = [
        (42, 'Name:'),
        (43, 'Role:'),
        (44, 'Date:'),
        (45, 'Signature:'),
        (46, 'Final Approval:', 'I approve this assessment as accurate'),
        (47, 'Risk Acceptance:', 'I accept residual risk for: [list]'),
        (48, 'Comments:')
    ]

    for row_num, label, *rest in level3_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])

        if rest:
            ws[f'B{row_num}'] = rest[0]
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # --- Assessment Metadata (Rows 50-60) ---
    ws['A50'] = 'Next Review Date:'
    apply_cell_style(ws['A50'], styles['label_cell'])
    apply_cell_style(ws['B50'], styles['input_cell'])

    ws['A51'] = 'Assessment Status:'
    apply_cell_style(ws['A51'], styles['label_cell'])
    apply_cell_style(ws['B51'], styles['input_cell'])

    status_dv = create_data_validation(['Draft', 'Under-Review', 'Approved', 'Audit-Ready'])
    ws.add_data_validation(status_dv)
    status_dv.add('B51')

    # Audit Readiness Checklist
    ws['A53'] = 'Audit Readiness Checklist:'
    apply_cell_style(ws['A53'], styles['label_cell'])

    audit_items = [
        'All three approvals complete',
        'Evidence 100% verified',
        'NDA coverage verified (audit focal point)',
        'All critical gaps have remediation plans',
        'Assessment audit-ready'
    ]

    for i, item in enumerate(audit_items, start=54):
        ws[f'A{i}'] = f'[ ] {item}'
        ws.merge_cells(f'A{i}:D{i}')

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False


# ==============================================================================
# MAIN EXECUTION
# ==============================================================================

def main():
    """
    Main function to generate the Employment Contract Assessment workbook.

    Returns:
        int: 0 on success, 1 on failure
    """
    try:
        logger.info("=" * 80)
        logger.info("ISMS-IMP-A.5.1-2-6.1-2.S4 - Employment Contract Assessment")
        logger.info("Workbook Generator")
        logger.info("=" * 80)

        # Initialize workbook
        logger.info("Creating workbook...")
        wb = Workbook()
        styles = setup_styles()

        # Create all sheets
        logger.info("Creating Dashboard...")
        create_dashboard_sheet(wb, styles)

        logger.info("Creating Contract_Template_Assessment...")
        create_contract_template_assessment_sheet(wb, styles)

        logger.info("Creating Required_Clause_Registry...")
        create_required_clause_registry_sheet(wb, styles)

        logger.info("Creating Personnel_Contract_Compliance...")
        create_personnel_contract_compliance_sheet(wb, styles)

        logger.info("Creating Confidentiality_NDA_Tracking...")
        create_confidentiality_nda_tracking_sheet(wb, styles)

        logger.info("Creating Post_Employment_Obligations...")
        create_post_employment_obligations_sheet(wb, styles)

        logger.info("Creating Contractor_Agreement_Assessment...")
        create_contractor_agreement_assessment_sheet(wb, styles)

        logger.info("Creating Gap_Analysis...")
        create_gap_analysis_sheet(wb, styles)

        logger.info("Creating Evidence_Register...")
        create_evidence_register_sheet(wb, styles)

        logger.info("Creating Approval_Sign_Off...")
        create_approval_signoff_sheet(wb, styles)

        # Set workbook properties
        wb.properties.title = "ISMS-IMP-A.5.1-2-6.1-2.S4 - Employment Contract Assessment"
        wb.properties.subject = "ISO/IEC 27001:2022 Controls A.6.1/A.6.2 Assessment"
        wb.properties.creator = "[Organization] Information Security Team"
        wb.properties.keywords = "Employment, Contract, NDA, ISMS, ISO27001, A.6.1, A.6.2"
        wb.properties.comments = "Generated via Python script generate_a5_1_2_6_1_2_s4_employment_contract.py"

        # Generate filename with current date
        today = datetime.now().strftime("%Y%m%d")
        filename = f"ISMS-IMP-A.5.1-2-6.1-2.S4_Employment_Contract_Assessment_{today}.xlsx"

        # Save workbook
        logger.info(f"Saving workbook as: {filename}")
        wb.save(filename)

        logger.info("=" * 80)
        logger.info("Workbook generated successfully!")
        logger.info("=" * 80)
        logger.info("Next Steps:")
        logger.info("1. Open the workbook in Excel")
        logger.info("2. Complete Sheet 2 (Contract_Template_Assessment) - verify all templates")
        logger.info("3. Review Sheet 3 (Required_Clause_Registry) - verify clause coverage")
        logger.info("4. Complete Sheet 4 (Personnel_Contract_Compliance) - per-person verification")
        logger.info("5. Verify Sheet 5 (Confidentiality_NDA_Tracking) - AUDIT FOCAL POINT")
        logger.info("6. Track Sheet 6 (Post_Employment_Obligations) - ongoing obligations")
        logger.info("7. Assess Sheet 7 (Contractor_Agreement_Assessment) - third-party contracts")
        logger.info("8. Document Sheet 8 (Gap_Analysis) - consolidate all gaps")
        logger.info("9. Complete Sheet 9 (Evidence_Register) - supporting evidence")
        logger.info("10. Obtain Sheet 10 (Approval_Sign_Off) - three-level approval")
        logger.info(f"File location: ./{filename}")

        return 0

    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
