#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.1-2-6.1-2.S3 - Screening & Vetting Assessment Workbook Generator
================================================================================

Purpose:
    Generate Excel assessment workbook for ISO/IEC 27001:2022 Control A.6.1
    (Screening) as part of stacked control framework A.5.1 + A.5.2 + A.6.1 + A.6.2
    (Secure Employment and Roles).

Control Alignment:
    - ISO/IEC 27001:2022 Annex A.6.1: Screening
    - ISMS-POL-A.5.1-2-6.1-2, Section 6: Screening Requirements

Assessment Focus:
    - Pre-employment screening compliance
    - Screening levels matched to role sensitivity
    - Continuous/periodic re-screening
    - Swiss FADP compliance (consent, data retention)
    - Screening provider governance

Workbook Structure:
    10 sheets total:
    1. Dashboard - Executive summary, auto-calculated metrics
    2. Screening_Process_Assessment - Screening process documentation verification
    3. Screening_Level_Matrix - Role tier to screening level mapping
    4. Personnel_Screening_Registry - Master registry: all personnel screening status
    5. Screening_Compliance_Verif - Per-person compliance gap identification
    6. Continuous_Screening_Assessment - Re-screening schedule and compliance
    7. Legal_Compliance_Review - FADP/GDPR screening legality verification
    8. Gap_Analysis - All gaps consolidated with risk and remediation
    9. Evidence_Register - Supporting evidence documentation
   10. Approval_Sign_Off - Three-level approval workflow

Dependencies:
    - openpyxl >= 3.0.0

Usage:
    python generate_a5_1_2_6_1_2_s3_screening_vetting.py

Output:
    ISMS-IMP-A.5.1-2-6.1-2.S3_Screening_Vetting_{YYYYMMDD}.xlsx

Technical Specification:
    See ISMS-IMP-A.5.1-2-6.1-2.S3-Screening-Vetting-Assessment.md Part II

Quality Standards:
    - Follows A.8.23/A.8.24 reference implementation patterns
    - Comprehensive documentation and inline comments
    - Consistent styling (Yellow input, Light Blue calculated, Gray labels)
    - Data validation on all appropriate cells
    - Conditional formatting for compliance statuses
    - Sheet protection with selective unlocking
    - Named ranges for cross-sheet references
    - Formula-based auto-calculations
    - Professional formatting for executive presentation

Author:               [Organization] ISMS Implementation Team
Version: 1.0
Date:                 [Date to be set]
License: Internal Use Only

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta

# =============================================================================
# Third-Party Imports
# =============================================================================
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)




# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.1-2-6.1-2.S3"
WORKBOOK_NAME = "Screening Vetting Assessment Workbook"
CONTROL_ID = "A.5.1-2-6.1-2"
CONTROL_NAME = "Information Security Policies and Organization"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ==============================================================================
# STYLE DEFINITIONS
# ==============================================================================

def setup_styles():
    """
    Define consistent cell styles used throughout the workbook.

    Returns:
        dict: Dictionary of style objects for different cell types
    """
    styles = {}

    # Header styles
    styles['header_main'] = {
        'font': Font(name='Calibri', size=18, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['header_sub'] = {
        'font': Font(name='Calibri', size=12, bold=False, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['section_header'] = {
        'font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['column_header'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    # Data cell styles
    styles['input_cell'] = {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),  # Yellow
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    styles['calculated_cell'] = {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'),  # Light Blue
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    styles['label_cell'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),  # Gray
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    # Status fills (for conditional formatting reference)
    styles['status_compliant'] = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
    styles['status_partial'] = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')  # Yellow
    styles['status_non_compliant'] = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red

    return styles


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def apply_cell_style(cell, style_dict):
    """
    Apply a style dictionary to a cell.

    Args:
        cell: openpyxl cell object
        style_dict: Dictionary containing style attributes
    """
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']


def set_column_widths(ws, widths_dict):
    """
    Set column widths for a worksheet.

    Args:
        ws: Worksheet object
        widths_dict: Dictionary mapping column letters to widths
    """
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width


def create_data_validation(formula_or_list, allow_blank=True, error_title="Invalid Entry",
                           error_message="Please select from dropdown"):
    """
    Create a data validation object for dropdown lists.

    Args:
        formula_or_list: Either a string formula or list of values
        allow_blank: Whether to allow blank cells
        error_title: Title for error message
        error_message: Error message text

    Returns:
        DataValidation object
    """
    if isinstance(formula_or_list, list):
        dv = DataValidation(type="list", formula1=f'"{",".join(formula_or_list)}"',
                           allow_blank=allow_blank)
    else:
        dv = DataValidation(type="list", formula1=formula_or_list, allow_blank=allow_blank)

    dv.error = error_message
    dv.errorTitle = error_title
    dv.showErrorMessage = True
    return dv


def merge_and_style(ws, range_string, value, style_dict):
    """
    Merge cells and apply style.

    Args:
        ws: Worksheet object
        range_string: Range to merge (e.g., 'A1:K1')
        value: Value to write
        style_dict: Style dictionary to apply
    """
    ws.merge_cells(range_string)
    cell = ws[range_string.split(':')[0]]
    cell.value = value
    apply_cell_style(cell, style_dict)


# ==============================================================================
# SHEET CREATION FUNCTIONS
# ==============================================================================

def create_dashboard_sheet(wb, styles):
    """
    Sheet 1: Dashboard (Executive Summary)

    Auto-calculated metrics pulling from other sheets.
    Read-only for users.
    """
    ws = wb.active
    ws.title = "Dashboard"

    # --- Header Section ---
    merge_and_style(ws, 'A1:K1',
                   'ISMS-IMP-A.5.1-2-6.1-2.S3 - Screening & Vetting Assessment Dashboard',
                   styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:K2',
                   'ISO/IEC 27001:2022 Control A.6.1 - Screening',
                   styles['header_sub'])
    ws.row_dimensions[2].height = 30

    # --- Document Information Block (Rows 4-15) ---
    ws['A4'] = 'Document ID:'
    apply_cell_style(ws['A4'], styles['label_cell'])
    ws.merge_cells('B4:C4')
    ws['B4'] = 'ISMS-IMP-A.5.1-2-6.1-2.S3'

    info_rows = [
        (5, 'Related Policy:', 'ISMS-POL-A.5.1-2-6.1-2, Section 6'),
        (6, 'ISO Control:', 'A.6.1 (Screening)'),
        (7, 'Assessment Period:', None),  # User input
        (8, 'Assessment Date:', '=TODAY()'),  # Formula
        (9, 'Assessor Name:', None),  # User input
        (10, 'Assessor Role:', None),  # User input
        (11, 'Organization:', None),  # User input
        (12, 'Review Cycle:', 'Annual'),
        (13, 'Last Updated:', '=NOW()'),  # Formula
        (14, 'Assessment Status:', None),  # Dropdown
        (15, 'Next Review Date:', None)  # User input
    ]

    for row_num, label, value in info_rows:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        ws.merge_cells(f'B{row_num}:C{row_num}')

        if value:
            ws[f'B{row_num}'] = value
            if value.startswith('='):
                apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
            else:
                apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # Status dropdown for row 14
    status_dv = create_data_validation(['Draft', 'Under-Review', 'Approved', 'Audit-Ready'])
    ws.add_data_validation(status_dv)
    status_dv.add(ws['B14'])

    # --- Overall Compliance Summary (Rows 17-30) ---
    merge_and_style(ws, 'A17:K17', 'OVERALL SCREENING COMPLIANCE', styles['section_header'])
    ws.row_dimensions[17].height = 35

    # Compliance Scorecard (Rows 19-26)
    scorecard = [
        (19, 'Overall Compliance Score', '=AVERAGE(C32:C37)', '90%'),
        (20, 'Total Personnel Assessed', '=COUNTA(Personnel_Screening_Registry!A:A)-4', 'N/A'),
        (21, 'Personnel Compliant', '=COUNTIF(Screening_Compliance_Verif!M:M,"Compliant")', '100%'),
        (22, 'Personnel with Gaps', '=B20-B21', '0'),
        (23, 'Critical Gaps', '=COUNTIF(Gap_Analysis!F:F,"Critical")', '0'),
        (24, 'High Priority Gaps', '=COUNTIF(Gap_Analysis!F:F,"High")', '0'),
        (25, 'Medium Priority Gaps', '=COUNTIF(Gap_Analysis!F:F,"Medium")', 'N/A'),
        (26, 'Low Priority Gaps', '=COUNTIF(Gap_Analysis!F:F,"Low")', 'N/A')
    ]

    for row_num, label, formula, target in scorecard:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])

        ws[f'B{row_num}'] = formula
        apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        if row_num == 19:
            ws[f'B{row_num}'].number_format = '0.00%'

        ws[f'C{row_num}'] = target
        apply_cell_style(ws[f'C{row_num}'], styles['calculated_cell'])

        ws[f'D{row_num}'] = '=IF(B{0}>=C{0},"On Target",IF(B{0}>=C{0}*0.9,"Close","Below Target"))'.format(row_num)
        apply_cell_style(ws[f'D{row_num}'], styles['calculated_cell'])

    # --- Domain Compliance Breakdown (Rows 28-38) ---
    merge_and_style(ws, 'A28:K28', 'DOMAIN COMPLIANCE BREAKDOWN', styles['section_header'])

    # Column headers for domain table
    for col, header in zip(['A', 'B', 'C', 'D', 'E'],
                           ['Domain', 'Weight', 'Score (%)', 'Weighted Score', 'Status']):
        ws[f'{col}30'] = header
        apply_cell_style(ws[f'{col}30'], styles['column_header'])

    domains = [
        (32, 'Screening Process Documentation', '20%',
         '=COUNTIF(Screening_Process_Assessment!I:I,"Compliant")/MAX(1,COUNTA(Screening_Process_Assessment!A:A)-4)*100', '=C32*0.20'),
        (33, 'Screening Level Mapping', '15%',
         '=COUNTIF(Screening_Level_Matrix!I:I,"Yes")/MAX(1,COUNTA(Screening_Level_Matrix!A:A)-4)*100', '=C33*0.15'),
        (34, 'Personnel Screening Compliance', '25%',
         '=COUNTIF(Screening_Compliance_Verif!M:M,"Compliant")/MAX(1,COUNTA(Screening_Compliance_Verif!A:A)-4)*100', '=C34*0.25'),
        (35, 'Continuous Screening', '15%',
         '=COUNTIF(Continuous_Screening_Assessment!L:L,"Compliant")/MAX(1,COUNTA(Continuous_Screening_Assessment!A:A)-4)*100', '=C35*0.15'),
        (36, 'Legal Compliance (FADP/GDPR)', '20%',
         '=COUNTIF(Legal_Compliance_Review!I:I,"Compliant")/MAX(1,COUNTA(Legal_Compliance_Review!A:A)-4)*100', '=C36*0.20'),
        (37, 'Screening Provider Governance', '5%',
         '=COUNTIF(Screening_Process_Assessment!J:J,"Yes")/MAX(1,COUNTA(Screening_Process_Assessment!A:A)-4)*100', '=C37*0.05')
    ]

    for row_num, domain, weight, score_formula, weighted_formula in domains:
        ws[f'A{row_num}'] = domain
        ws[f'B{row_num}'] = weight
        ws[f'C{row_num}'] = score_formula
        ws[f'C{row_num}'].number_format = '0.00%'
        ws[f'D{row_num}'] = weighted_formula
        ws[f'D{row_num}'].number_format = '0.00%'
        ws[f'E{row_num}'] = f'=IF(C{row_num}>=0.9,"Pass",IF(C{row_num}>=0.7,"Review","Fail"))'

    # Total row
    ws['A38'] = 'TOTAL WEIGHTED SCORE'
    apply_cell_style(ws['A38'], styles['label_cell'])
    ws['B38'] = '100%'
    ws['C38'] = '=SUM(D32:D37)'
    ws['C38'].number_format = '0.00%'
    ws['C38'].font = Font(name='Calibri', size=10, bold=True)

    # Set column widths
    set_column_widths(ws, {'A': 40, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 5, 'G': 5, 'H': 5, 'I': 5, 'J': 5, 'K': 5})

    # Protect sheet (allow sorting/filtering)
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_screening_process_assessment_sheet(wb, styles):
    """
    Sheet 2: Screening_Process_Assessment

    Screening process documentation verification.
    """
    ws = wb.create_sheet("Screening_Process_Assessment")

    # --- Header ---
    merge_and_style(ws, 'A1:L1', 'SCREENING PROCESS ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:L2',
                   'Verification of screening process documentation and implementation',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # --- Column Headers (Row 4) ---
    headers = [
        ('A', 'Process_ID', 15),
        ('B', 'Process_Name', 35),
        ('C', 'Process_Owner', 25),
        ('D', 'Documented', 12),
        ('E', 'Document_Location', 35),
        ('F', 'Last_Review_Date', 15),
        ('G', 'Process_Implemented', 15),
        ('H', 'Implementation_Evidence', 35),
        ('I', 'Compliance_Status', 15),
        ('J', 'Provider_Governed', 12),
        ('K', 'Gap_Description', 40),
        ('L', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)
    ws.row_dimensions[4].height = 30

    # Pre-populate process rows (5-24, 20 processes)
    processes = [
        ('SP-001', 'Pre-Employment Background Check Process'),
        ('SP-002', 'Identity Verification Process'),
        ('SP-003', 'Employment History Verification'),
        ('SP-004', 'Education Verification Process'),
        ('SP-005', 'Professional Certification Verification'),
        ('SP-006', 'Criminal Record Check Process'),
        ('SP-007', 'Credit Check Process (where applicable)'),
        ('SP-008', 'Reference Check Process'),
        ('SP-009', 'Right-to-Work Verification'),
        ('SP-010', 'Security Clearance Process'),
        ('SP-011', 'Contractor Screening Process'),
        ('SP-012', 'Third-Party Personnel Screening'),
        ('SP-013', 'Re-Screening Trigger Process'),
        ('SP-014', 'Screening Results Documentation'),
        ('SP-015', 'Screening Exception Handling'),
        ('SP-016', 'Screening Provider Management'),
        ('SP-017', 'Consent Management Process'),
        ('SP-018', 'Data Retention Process'),
        ('SP-019', 'Screening Appeals Process'),
        ('SP-020', 'Screening Audit Process')
    ]

    for idx, (proc_id, proc_name) in enumerate(processes, start=5):
        ws[f'A{idx}'] = proc_id
        apply_cell_style(ws[f'A{idx}'], styles['calculated_cell'])
        ws[f'B{idx}'] = proc_name
        apply_cell_style(ws[f'B{idx}'], styles['calculated_cell'])

        # Manual entry cells
        for col in ['C', 'E', 'H', 'K', 'L']:
            apply_cell_style(ws[f'{col}{idx}'], styles['input_cell'])

        # Cells with dropdowns
        for col in ['D', 'F', 'G', 'I', 'J']:
            apply_cell_style(ws[f'{col}{idx}'], styles['input_cell'])

    # Additional rows for custom processes (25-44)
    for row in range(25, 45):
        for col_letter in [c[0] for c in headers]:
            apply_cell_style(ws[f'{col_letter}{row}'], styles['input_cell'])

    # --- Data Validation ---
    documented_dv = create_data_validation(['Yes', 'Partial', 'No'])
    ws.add_data_validation(documented_dv)
    documented_dv.add('D5:D44')

    implemented_dv = create_data_validation(['Yes', 'Partial', 'No', 'Not-Required'])
    ws.add_data_validation(implemented_dv)
    implemented_dv.add('G5:G44')

    compliance_dv = create_data_validation(['Compliant', 'Partial', 'Non-Compliant'])
    ws.add_data_validation(compliance_dv)
    compliance_dv.add('I5:I44')

    provider_dv = create_data_validation(['Yes', 'No', 'N/A'])
    ws.add_data_validation(provider_dv)
    provider_dv.add('J5:J44')

    # Date formatting
    for row in range(5, 45):
        ws[f'F{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes at A5
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_screening_level_matrix_sheet(wb, styles):
    """
    Sheet 3: Screening_Level_Matrix

    Role tier to screening level mapping.
    """
    ws = wb.create_sheet("Screening_Level_Matrix")

    # --- Header ---
    merge_and_style(ws, 'A1:K1', 'SCREENING LEVEL MATRIX', styles['header_main'])
    merge_and_style(ws, 'A2:K2',
                   'Role sensitivity tier to screening level mapping and requirements',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # --- Column Headers (Row 4) ---
    headers = [
        ('A', 'Role_Tier', 15),
        ('B', 'Role_Category', 30),
        ('C', 'Example_Roles', 40),
        ('D', 'Required_Screening_Level', 20),
        ('E', 'Identity_Check', 12),
        ('F', 'Employment_History', 12),
        ('G', 'Criminal_Check', 12),
        ('H', 'Credit_Check', 12),
        ('I', 'Mapping_Documented', 12),
        ('J', 'Gap_Description', 40),
        ('K', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)
    ws.row_dimensions[4].height = 30

    # Pre-populate tier rows
    tiers = [
        ('Tier-1', 'Executive/C-Suite', 'CEO, CFO, CTO, CISO', 'Enhanced'),
        ('Tier-2', 'Senior Management', 'Directors, VPs, Department Heads', 'Enhanced'),
        ('Tier-3', 'Information Security', 'Security Analysts, SOC Staff, Admins', 'Enhanced'),
        ('Tier-4', 'IT Privileged Access', 'System Admins, DBAs, Network Engineers', 'Enhanced'),
        ('Tier-5', 'Finance/Treasury', 'Accountants, Controllers, Treasury Staff', 'Enhanced'),
        ('Tier-6', 'HR/Legal', 'HR Managers, Legal Counsel, Recruiters', 'Standard-Plus'),
        ('Tier-7', 'Customer Data Access', 'Customer Service, Sales, Support', 'Standard'),
        ('Tier-8', 'General Staff', 'Administrative, Operations', 'Standard'),
        ('Tier-9', 'Contractors - High Risk', 'IT Contractors, Security Consultants', 'Enhanced'),
        ('Tier-10', 'Contractors - Standard', 'General Contractors', 'Standard'),
        ('Tier-11', 'Temporary Staff', 'Temps, Interns', 'Basic'),
        ('Tier-12', 'Third-Party On-Site', 'Vendors with physical access', 'Standard')
    ]

    for idx, (tier, category, examples, level) in enumerate(tiers, start=5):
        ws[f'A{idx}'] = tier
        apply_cell_style(ws[f'A{idx}'], styles['calculated_cell'])
        ws[f'B{idx}'] = category
        apply_cell_style(ws[f'B{idx}'], styles['calculated_cell'])
        ws[f'C{idx}'] = examples
        apply_cell_style(ws[f'C{idx}'], styles['calculated_cell'])
        ws[f'D{idx}'] = level
        apply_cell_style(ws[f'D{idx}'], styles['input_cell'])

        # Screening requirement cells
        for col in ['E', 'F', 'G', 'H', 'I']:
            apply_cell_style(ws[f'{col}{idx}'], styles['input_cell'])

        # Gap and notes
        for col in ['J', 'K']:
            apply_cell_style(ws[f'{col}{idx}'], styles['input_cell'])

    # Additional rows for custom tiers (17-30)
    for row in range(17, 31):
        for col_letter in [c[0] for c in headers]:
            apply_cell_style(ws[f'{col_letter}{row}'], styles['input_cell'])

    # --- Data Validation ---
    screening_level_dv = create_data_validation(['Enhanced', 'Standard-Plus', 'Standard', 'Basic', 'None'])
    ws.add_data_validation(screening_level_dv)
    screening_level_dv.add('D5:D30')

    yes_no_req_dv = create_data_validation(['Required', 'Optional', 'Not-Required'])
    ws.add_data_validation(yes_no_req_dv)
    yes_no_req_dv.add('E5:H30')

    documented_dv = create_data_validation(['Yes', 'No'])
    ws.add_data_validation(documented_dv)
    documented_dv.add('I5:I30')

    # Define named range for Role_Tier_List
    tier_range = DefinedName(name='Role_Tier_List', attr_text="Screening_Level_Matrix!$A$5:$A$30")
    wb.defined_names.add(tier_range)

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_personnel_screening_registry_sheet(wb, styles):
    """
    Sheet 4: Personnel_Screening_Registry

    Master registry: all personnel screening status.
    """
    ws = wb.create_sheet("Personnel_Screening_Registry")

    # --- Header ---
    merge_and_style(ws, 'A1:R1', 'PERSONNEL SCREENING REGISTRY', styles['header_main'])
    merge_and_style(ws, 'A2:R2',
                   'Master registry of all personnel with screening status and dates',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # --- Column Headers (Row 4) ---
    headers = [
        ('A', 'Employee_ID', 15),
        ('B', 'Full_Name', 25),
        ('C', 'Department', 20),
        ('D', 'Job_Title', 25),
        ('E', 'Role_Tier', 12),
        ('F', 'Employment_Type', 15),
        ('G', 'Start_Date', 12),
        ('H', 'Required_Screening_Level', 15),
        ('I', 'Actual_Screening_Level', 15),
        ('J', 'Screening_Complete', 12),
        ('K', 'Screening_Date', 12),
        ('L', 'Screening_Expiry', 12),
        ('M', 'Screening_Provider', 20),
        ('N', 'Consent_Obtained', 12),
        ('O', 'Consent_Date', 12),
        ('P', 'Screening_Status', 15),
        ('Q', 'Gap_Identified', 15),
        ('R', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)
    ws.row_dimensions[4].height = 30

    # --- Data Rows (5-254, 250 personnel) ---
    for row in range(5, 255):
        # All input cells
        for col_letter in [c[0] for c in headers]:
            cell = ws[f'{col_letter}{row}']
            apply_cell_style(cell, styles['input_cell'])

        # Auto-calculated: Screening_Status (P)
        ws[f'P{row}'] = (f'=IF(A{row}="","",IF(AND(J{row}="Yes",L{row}>=TODAY()),"Valid",'
                        f'IF(AND(J{row}="Yes",L{row}<TODAY()),"Expired",'
                        f'IF(J{row}="In-Progress","Pending","Not-Screened"))))')
        apply_cell_style(ws[f'P{row}'], styles['calculated_cell'])

        # Auto-calculated: Gap_Identified (Q)
        ws[f'Q{row}'] = (f'=IF(A{row}="","",IF(AND(H{row}=I{row},J{row}="Yes",L{row}>=TODAY()),"No-Gap",'
                        f'IF(OR(J{row}="No",L{row}<TODAY()),"Critical-Gap",'
                        f'IF(H{row}<>I{row},"Level-Gap","Minor-Gap"))))')
        apply_cell_style(ws[f'Q{row}'], styles['calculated_cell'])

    # --- Data Validation ---
    tier_dv = create_data_validation(['Tier-1', 'Tier-2', 'Tier-3', 'Tier-4', 'Tier-5', 'Tier-6',
                                      'Tier-7', 'Tier-8', 'Tier-9', 'Tier-10', 'Tier-11', 'Tier-12'])
    ws.add_data_validation(tier_dv)
    tier_dv.add('E5:E254')

    emp_type_dv = create_data_validation(['Full-Time', 'Part-Time', 'Contractor', 'Temporary', 'Intern', 'Third-Party'])
    ws.add_data_validation(emp_type_dv)
    emp_type_dv.add('F5:F254')

    screening_level_dv = create_data_validation(['Enhanced', 'Standard-Plus', 'Standard', 'Basic', 'None'])
    ws.add_data_validation(screening_level_dv)
    screening_level_dv.add('H5:I254')

    complete_dv = create_data_validation(['Yes', 'No', 'In-Progress', 'Pending'])
    ws.add_data_validation(complete_dv)
    complete_dv.add('J5:J254')

    consent_dv = create_data_validation(['Yes', 'No', 'Pending'])
    ws.add_data_validation(consent_dv)
    consent_dv.add('N5:N254')

    # Date formatting
    for col in ['G', 'K', 'L', 'O']:
        for row in range(5, 255):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # Define named range for Employee_ID_List
    emp_id_range = DefinedName(name='Employee_ID_List', attr_text="Personnel_Screening_Registry!$A$5:$A$254")
    wb.defined_names.add(emp_id_range)

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_screening_compliance_verification_sheet(wb, styles):
    """
    Sheet 5: Screening_Compliance_Verif

    Per-person compliance gap identification.
    """
    ws = wb.create_sheet("Screening_Compliance_Verif")

    # --- Header ---
    merge_and_style(ws, 'A1:O1', 'SCREENING COMPLIANCE VERIFICATION', styles['header_main'])
    merge_and_style(ws, 'A2:O2',
                   'Per-person screening compliance verification and gap identification',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # --- Column Headers (Row 4) ---
    headers = [
        ('A', 'Employee_ID', 15),
        ('B', 'Full_Name', 25),
        ('C', 'Role_Tier', 12),
        ('D', 'Identity_Verified', 12),
        ('E', 'Employment_History_Checked', 12),
        ('F', 'Education_Verified', 12),
        ('G', 'Criminal_Check_Complete', 12),
        ('H', 'Credit_Check_Complete', 12),
        ('I', 'References_Checked', 12),
        ('J', 'Right_to_Work_Verified', 12),
        ('K', 'All_Required_Checks_Complete', 15),
        ('L', 'Screening_Level_Appropriate', 15),
        ('M', 'Compliance_Status', 15),
        ('N', 'Gap_Description', 40),
        ('O', 'Evidence_Reference', 30)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)
    ws.row_dimensions[4].height = 30

    # --- Data Rows (5-254) with Auto-Population Formulas ---
    for row in range(5, 255):
        # Auto-populate from Personnel_Screening_Registry
        ws[f'A{row}'] = f'=IF(ROW()-4<=COUNTA(Personnel_Screening_Registry!$A:$A)-4,INDEX(Personnel_Screening_Registry!$A:$A,ROW()-3),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Personnel_Screening_Registry!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])

        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Personnel_Screening_Registry!$A:$E,5,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])

        # Manual entry cells (D-J, N-O)
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'N', 'O']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # All_Required_Checks_Complete (K) - auto-calculated
        ws[f'K{row}'] = (f'=IF(A{row}="","",IF(AND(D{row}="Yes",E{row}="Yes",J{row}="Yes"),"Yes","No"))')
        apply_cell_style(ws[f'K{row}'], styles['calculated_cell'])

        # Screening_Level_Appropriate (L) - auto-calculated
        ws[f'L{row}'] = (f'=IF(A{row}="","",IF(VLOOKUP(A{row},Personnel_Screening_Registry!$A:$I,8,FALSE)='
                        f'VLOOKUP(A{row},Personnel_Screening_Registry!$A:$I,9,FALSE),"Yes","No"))')
        apply_cell_style(ws[f'L{row}'], styles['calculated_cell'])

        # Compliance_Status (M) - auto-calculated
        ws[f'M{row}'] = (f'=IF(A{row}="","",IF(AND(K{row}="Yes",L{row}="Yes"),"Compliant",'
                        f'IF(OR(K{row}="No",L{row}="No"),"Non-Compliant","Partial")))')
        apply_cell_style(ws[f'M{row}'], styles['calculated_cell'])

    # --- Data Validation ---
    check_dv = create_data_validation(['Yes', 'No', 'N/A', 'Pending'])
    ws.add_data_validation(check_dv)
    check_dv.add('D5:J254')

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_continuous_screening_assessment_sheet(wb, styles):
    """
    Sheet 6: Continuous_Screening_Assessment

    Re-screening schedule and compliance.
    """
    ws = wb.create_sheet("Continuous_Screening_Assessment")

    # --- Header ---
    merge_and_style(ws, 'A1:N1', 'CONTINUOUS SCREENING ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:N2',
                   'Re-screening schedule tracking and continuous monitoring compliance',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # --- Column Headers (Row 4) ---
    headers = [
        ('A', 'Employee_ID', 15),
        ('B', 'Full_Name', 25),
        ('C', 'Role_Tier', 12),
        ('D', 'Initial_Screening_Date', 15),
        ('E', 'Re_Screening_Interval', 15),
        ('F', 'Last_Re_Screening_Date', 15),
        ('G', 'Next_Re_Screening_Due', 15),
        ('H', 'Re_Screening_Status', 15),
        ('I', 'Trigger_Event_Occurred', 15),
        ('J', 'Trigger_Event_Details', 30),
        ('K', 'Continuous_Monitoring_Active', 15),
        ('L', 'Compliance_Status', 15),
        ('M', 'Gap_Description', 40),
        ('N', 'Evidence_Reference', 30)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)
    ws.row_dimensions[4].height = 30

    # --- Data Rows (5-254) with Auto-Population Formulas ---
    for row in range(5, 255):
        # Auto-populate from Personnel_Screening_Registry
        ws[f'A{row}'] = f'=IF(ROW()-4<=COUNTA(Personnel_Screening_Registry!$A:$A)-4,INDEX(Personnel_Screening_Registry!$A:$A,ROW()-3),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Personnel_Screening_Registry!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])

        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Personnel_Screening_Registry!$A:$E,5,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])

        ws[f'D{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Personnel_Screening_Registry!$A:$K,11,FALSE),"")'
        apply_cell_style(ws[f'D{row}'], styles['calculated_cell'])
        ws[f'D{row}'].number_format = 'DD.MM.YYYY'

        # Manual entry cells
        for col in ['E', 'F', 'I', 'J', 'K', 'M', 'N']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # Next_Re_Screening_Due (G) - auto-calculated
        ws[f'G{row}'] = (f'=IF(OR(A{row}="",F{row}="",E{row}=""),"",IF(E{row}="Annual",F{row}+365,'
                        f'IF(E{row}="Biennial",F{row}+730,IF(E{row}="3-Year",F{row}+1095,F{row}+365))))')
        apply_cell_style(ws[f'G{row}'], styles['calculated_cell'])
        ws[f'G{row}'].number_format = 'DD.MM.YYYY'

        # Re_Screening_Status (H) - auto-calculated
        ws[f'H{row}'] = (f'=IF(A{row}="","",IF(G{row}="","Not-Scheduled",'
                        f'IF(G{row}>=TODAY(),"Current",IF(G{row}>=TODAY()-30,"Due-Soon","Overdue"))))')
        apply_cell_style(ws[f'H{row}'], styles['calculated_cell'])

        # Compliance_Status (L) - auto-calculated
        ws[f'L{row}'] = (f'=IF(A{row}="","",IF(AND(OR(H{row}="Current",H{row}="Due-Soon"),'
                        f'OR(I{row}="No",I{row}="N/A",AND(I{row}="Yes",K{row}="Yes"))),"Compliant",'
                        f'IF(H{row}="Overdue","Non-Compliant","Partial")))')
        apply_cell_style(ws[f'L{row}'], styles['calculated_cell'])

    # Date formatting
    for row in range(5, 255):
        ws[f'F{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    interval_dv = create_data_validation(['Annual', 'Biennial', '3-Year', 'Triggered-Only', 'N/A'])
    ws.add_data_validation(interval_dv)
    interval_dv.add('E5:E254')

    trigger_dv = create_data_validation(['Yes', 'No', 'N/A'])
    ws.add_data_validation(trigger_dv)
    trigger_dv.add('I5:I254')

    monitoring_dv = create_data_validation(['Yes', 'No', 'N/A'])
    ws.add_data_validation(monitoring_dv)
    monitoring_dv.add('K5:K254')

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_legal_compliance_review_sheet(wb, styles):
    """
    Sheet 7: Legal_Compliance_Review

    FADP/GDPR screening legality verification.
    """
    ws = wb.create_sheet("Legal_Compliance_Review")

    # --- Header ---
    merge_and_style(ws, 'A1:K1', 'LEGAL COMPLIANCE REVIEW', styles['header_main'])
    merge_and_style(ws, 'A2:K2',
                   'Swiss FADP and GDPR screening legality verification',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # --- Column Headers (Row 4) ---
    headers = [
        ('A', 'Requirement_ID', 15),
        ('B', 'Requirement_Category', 25),
        ('C', 'Requirement_Description', 50),
        ('D', 'Applicable_Regulation', 20),
        ('E', 'Implementation_Status', 15),
        ('F', 'Implementation_Evidence', 35),
        ('G', 'Last_Review_Date', 15),
        ('H', 'Responsible_Party', 25),
        ('I', 'Compliance_Status', 15),
        ('J', 'Gap_Description', 40),
        ('K', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)
    ws.row_dimensions[4].height = 30

    # Pre-populate legal requirements
    requirements = [
        ('LR-001', 'Consent', 'Written consent obtained before screening', 'FADP Art. 6'),
        ('LR-002', 'Consent', 'Consent is informed (purpose, scope, retention)', 'FADP Art. 6'),
        ('LR-003', 'Consent', 'Consent is freely given (not condition of employment)', 'FADP Art. 6'),
        ('LR-004', 'Consent', 'Consent withdrawal mechanism documented', 'FADP Art. 6'),
        ('LR-005', 'Purpose Limitation', 'Screening limited to job-relevant checks', 'FADP Art. 6'),
        ('LR-006', 'Purpose Limitation', 'No excessive data collection', 'FADP Art. 6'),
        ('LR-007', 'Data Minimization', 'Only necessary data collected', 'GDPR Art. 5'),
        ('LR-008', 'Data Minimization', 'Screening scope proportionate to role', 'FADP Art. 6'),
        ('LR-009', 'Data Retention', 'Retention period defined and documented', 'FADP Art. 6'),
        ('LR-010', 'Data Retention', 'Data deleted after retention period', 'FADP Art. 6'),
        ('LR-011', 'Data Retention', 'Unsuccessful applicant data deleted (max 3 months)', 'FADP Art. 6'),
        ('LR-012', 'Transparency', 'Candidate informed of screening results', 'FADP Art. 25'),
        ('LR-013', 'Transparency', 'Candidate can access their screening data', 'FADP Art. 25'),
        ('LR-014', 'Transparency', 'Right to rectification documented', 'FADP Art. 32'),
        ('LR-015', 'Third-Party', 'Screening provider DPA in place', 'FADP Art. 9'),
        ('LR-016', 'Third-Party', 'Provider data handling compliant', 'FADP Art. 9'),
        ('LR-017', 'Cross-Border', 'Cross-border transfer safeguards (if applicable)', 'FADP Art. 16-17'),
        ('LR-018', 'Criminal Records', 'Criminal check legally permitted for role', 'Swiss Criminal Records Act'),
        ('LR-019', 'Credit Check', 'Credit check legally permitted for role', 'FADP Art. 6'),
        ('LR-020', 'Appeals', 'Adverse decision appeal process documented', 'Best Practice')
    ]

    for idx, (req_id, category, description, regulation) in enumerate(requirements, start=5):
        ws[f'A{idx}'] = req_id
        apply_cell_style(ws[f'A{idx}'], styles['calculated_cell'])
        ws[f'B{idx}'] = category
        apply_cell_style(ws[f'B{idx}'], styles['calculated_cell'])
        ws[f'C{idx}'] = description
        apply_cell_style(ws[f'C{idx}'], styles['calculated_cell'])
        ws[f'D{idx}'] = regulation
        apply_cell_style(ws[f'D{idx}'], styles['calculated_cell'])

        # Manual entry cells
        for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K']:
            apply_cell_style(ws[f'{col}{idx}'], styles['input_cell'])

    # Additional rows for custom requirements (25-44)
    for row in range(25, 45):
        for col_letter in [c[0] for c in headers]:
            apply_cell_style(ws[f'{col_letter}{row}'], styles['input_cell'])

    # --- Data Validation ---
    implementation_dv = create_data_validation(['Implemented', 'Partial', 'Not-Implemented', 'N/A'])
    ws.add_data_validation(implementation_dv)
    implementation_dv.add('E5:E44')

    compliance_dv = create_data_validation(['Compliant', 'Partial', 'Non-Compliant', 'N/A'])
    ws.add_data_validation(compliance_dv)
    compliance_dv.add('I5:I44')

    # Date formatting
    for row in range(5, 45):
        ws[f'G{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_gap_analysis_sheet(wb, styles):
    """
    Sheet 8: Gap_Analysis

    Consolidated gaps from all assessment domains.
    """
    ws = wb.create_sheet("Gap_Analysis")

    # --- Header ---
    merge_and_style(ws, 'A1:P1', 'GAP ANALYSIS', styles['header_main'])
    merge_and_style(ws, 'A2:P2',
                   'Consolidated screening gaps from all assessment domains with risk levels and remediation plans',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # --- Column Headers (Row 4) ---
    headers = [
        ('A', 'Gap_ID', 12),
        ('B', 'Source_Sheet', 25),
        ('C', 'Employee_ID', 15),
        ('D', 'Gap_Category', 20),
        ('E', 'Gap_Description', 40),
        ('F', 'Risk_Level', 12),
        ('G', 'Impact_Assessment', 35),
        ('H', 'Affected_Stakeholders', 25),
        ('I', 'Remediation_Action', 40),
        ('J', 'Responsible_Party', 25),
        ('K', 'Target_Completion_Date', 15),
        ('L', 'Estimated_Effort', 15),
        ('M', 'Dependencies', 30),
        ('N', 'Status', 15),
        ('O', 'Completion_Evidence', 30),
        ('P', 'Risk_Acceptance', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)
    ws.row_dimensions[4].height = 30

    # --- Data Rows (5-154, 150 rows for gaps) ---
    for row in range(5, 155):
        # Gap_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}="","",TEXT(ROW()-4,"SCR-GAP-000"))'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # All other columns manual entry
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

    # Date formatting for Target_Completion_Date
    for row in range(5, 155):
        ws[f'K{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    source_dv = create_data_validation(['Screening_Process_Assessment', 'Screening_Level_Matrix',
                                        'Personnel_Screening_Registry', 'Screening_Compliance_Verif',
                                        'Continuous_Screening_Assessment', 'Legal_Compliance_Review'])
    ws.add_data_validation(source_dv)
    source_dv.add('B5:B154')

    category_dv = create_data_validation(['Process-Gap', 'Documentation-Gap', 'Screening-Level-Gap',
                                          'Personnel-Gap', 'Re-Screening-Gap', 'Legal-Gap',
                                          'Provider-Gap', 'Consent-Gap', 'Retention-Gap'])
    ws.add_data_validation(category_dv)
    category_dv.add('D5:D154')

    risk_level_dv = create_data_validation(['Critical', 'High', 'Medium', 'Low'])
    ws.add_data_validation(risk_level_dv)
    risk_level_dv.add('F5:F154')

    effort_dv = create_data_validation(['<1hr', '1-4hrs', '1day', '2-5days', '>1week'])
    ws.add_data_validation(effort_dv)
    effort_dv.add('L5:L154')

    status_dv = create_data_validation(['Not-Started', 'In-Progress', 'Blocked', 'Completed', 'Accepted-Risk'])
    ws.add_data_validation(status_dv)
    status_dv.add('N5:N154')

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_evidence_register_sheet(wb, styles):
    """
    Sheet 9: Evidence_Register

    Supporting evidence documentation.
    """
    ws = wb.create_sheet("Evidence_Register")

    # --- Header ---
    merge_and_style(ws, 'A1:K1', 'EVIDENCE REGISTER', styles['header_main'])
    merge_and_style(ws, 'A2:K2',
                   'Documentation of all supporting evidence for screening assessment',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # --- Column Headers (Row 4) ---
    headers = [
        ('A', 'Evidence_ID', 12),
        ('B', 'Evidence_Type', 25),
        ('C', 'Description', 40),
        ('D', 'Related_Employee_ID', 15),
        ('E', 'Related_Assessment_Sheet', 25),
        ('F', 'File_Location', 40),
        ('G', 'Date_Collected', 12),
        ('H', 'Collected_By', 25),
        ('I', 'Verification_Status', 15),
        ('J', 'Retention_Date', 12),
        ('K', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)
    ws.row_dimensions[4].height = 30

    # --- Data Rows (5-204, 200 evidence items) ---
    for row in range(5, 205):
        # Evidence_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}="","",TEXT(ROW()-4,"SCR-EVD-000"))'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # All other columns manual entry
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

    # Date formatting
    for row in range(5, 205):
        ws[f'G{row}'].number_format = 'DD.MM.YYYY'
        ws[f'J{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    evidence_type_dv = create_data_validation(['Consent-Form', 'Screening-Report', 'Background-Check-Result',
                                               'Criminal-Record-Check', 'Credit-Check-Report', 'Reference-Letter',
                                               'Education-Certificate', 'Employment-Verification', 'ID-Document',
                                               'Right-to-Work-Document', 'Provider-Contract', 'DPA-Agreement',
                                               'Process-Document', 'Policy-Document', 'Training-Record', 'Other'])
    ws.add_data_validation(evidence_type_dv)
    evidence_type_dv.add('B5:B204')

    sheet_dv = create_data_validation(['Screening_Process_Assessment', 'Screening_Level_Matrix',
                                       'Personnel_Screening_Registry', 'Screening_Compliance_Verif',
                                       'Continuous_Screening_Assessment', 'Legal_Compliance_Review', 'Gap_Analysis'])
    ws.add_data_validation(sheet_dv)
    sheet_dv.add('E5:E204')

    verification_dv = create_data_validation(['Verified', 'Pending', 'Not-Verified', 'Expired'])
    ws.add_data_validation(verification_dv)
    verification_dv.add('I5:I204')

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_approval_signoff_sheet(wb, styles):
    """
    Sheet 10: Approval_Sign_Off

    Three-level approval workflow.
    """
    ws = wb.create_sheet("Approval_Sign_Off")

    # --- Header ---
    merge_and_style(ws, 'A1:D1', 'APPROVAL & SIGN-OFF', styles['header_main'])
    ws.row_dimensions[1].height = 30

    # --- Assessment Summary (Rows 3-16) ---
    ws['A3'] = 'ASSESSMENT SUMMARY'
    apply_cell_style(ws['A3'], styles['section_header'])
    ws.merge_cells('A3:D3')

    summary_items = [
        (5, 'Document:', 'ISMS-IMP-A.5.1-2-6.1-2.S3'),
        (6, 'Assessment Period:', '=Dashboard!B7'),
        (7, 'Overall Compliance Score:', '=Dashboard!B19'),
        (8, 'Total Personnel Assessed:', '=Dashboard!B20'),
        (9, 'Personnel Compliant:', '=Dashboard!B21'),
        (10, 'Personnel with Gaps:', '=Dashboard!B22'),
        (11, 'Critical Gaps:', '=Dashboard!B23'),
        (12, 'High Priority Gaps:', '=Dashboard!B24'),
        (13, 'Legal Compliance:', '=COUNTIF(Legal_Compliance_Review!I:I,"Compliant")/MAX(1,COUNTA(Legal_Compliance_Review!A:A)-4)*100&"%"')
    ]

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30

    for row_num, label, value in summary_items:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])

        ws[f'B{row_num}'] = value
        if value.startswith('='):
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])

    # --- Level 1 Approval (Rows 18-26) ---
    ws['A18'] = 'LEVEL 1: PREPARED BY (HR/Security Analyst)'
    apply_cell_style(ws['A18'], styles['section_header'])
    ws.merge_cells('A18:D18')

    level1_fields = [
        (19, 'Name:'),
        (20, 'Role:'),
        (21, 'Date:'),
        (22, 'Signature:'),
        (23, 'Certification:', 'I certify this assessment is complete and accurate'),
        (24, 'Comments:')
    ]

    for row_num, label, *rest in level1_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])

        if rest:
            ws[f'B{row_num}'] = rest[0]
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # --- Level 2 Approval (Rows 28-42) ---
    ws['A28'] = 'LEVEL 2: REVIEWED BY (HR Manager / Security Manager)'
    apply_cell_style(ws['A28'], styles['section_header'])
    ws.merge_cells('A28:D28')

    level2_fields = [
        (29, 'Name:'),
        (30, 'Role:'),
        (31, 'Date:'),
        (32, 'Signature:')
    ]

    for row_num, label in level2_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # Review checklist
    checklist_items = [
        'All personnel screened appropriately',
        'Screening levels match role sensitivity',
        'Re-screening schedule current',
        'FADP/GDPR compliance verified',
        'Consent documentation complete',
        'Screening provider governance verified',
        'Gap remediation plans appropriate',
        'Evidence sufficient for audit'
    ]

    ws['A34'] = 'Review Checklist:'
    apply_cell_style(ws['A34'], styles['label_cell'])

    for i, item in enumerate(checklist_items, start=35):
        ws[f'A{i}'] = f'[ ] {item}'
        ws.merge_cells(f'A{i}:D{i}')

    ws['A43'] = 'Comments:'
    apply_cell_style(ws['A43'], styles['label_cell'])
    apply_cell_style(ws['B43'], styles['input_cell'])

    # --- Level 3 Approval (Rows 45-56) ---
    ws['A45'] = 'LEVEL 3: APPROVED BY (CISO / DPO)'
    apply_cell_style(ws['A45'], styles['section_header'])
    ws.merge_cells('A45:D45')

    level3_fields = [
        (46, 'Name:'),
        (47, 'Role:'),
        (48, 'Date:'),
        (49, 'Signature:'),
        (50, 'Final Approval:', 'I approve this screening assessment as accurate'),
        (51, 'Risk Acceptance:', 'I accept residual screening risk for: [list]'),
        (52, 'Comments:')
    ]

    for row_num, label, *rest in level3_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])

        if rest:
            ws[f'B{row_num}'] = rest[0]
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # --- Assessment Metadata (Rows 55-61) ---
    ws['A55'] = 'Next Review Date:'
    apply_cell_style(ws['A55'], styles['label_cell'])
    apply_cell_style(ws['B55'], styles['input_cell'])

    ws['A56'] = 'Assessment Status:'
    apply_cell_style(ws['A56'], styles['label_cell'])
    apply_cell_style(ws['B56'], styles['input_cell'])

    status_dv = create_data_validation(['Draft', 'Under-Review', 'Approved', 'Audit-Ready'])
    ws.add_data_validation(status_dv)
    status_dv.add('B56')

    # Audit Readiness Checklist
    ws['A58'] = 'Audit Readiness Checklist:'
    apply_cell_style(ws['A58'], styles['label_cell'])

    audit_items = [
        'All three approvals complete',
        'Evidence 100% verified',
        'All critical gaps have remediation plans',
        'FADP/GDPR compliance verified',
        'Assessment audit-ready'
    ]

    for i, item in enumerate(audit_items, start=59):
        ws[f'A{i}'] = f'[ ] {item}'
        ws.merge_cells(f'A{i}:D{i}')

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False


# ==============================================================================
# MAIN EXECUTION
# ==============================================================================

def main():
    """
    Main function to generate the Screening & Vetting Assessment workbook.

    Returns:
        int: 0 on success, 1 on failure
    """
    try:
        logger.info("=" * 80)
        logger.info("ISMS-IMP-A.5.1-2-6.1-2.S3 - Screening & Vetting Assessment")
        logger.info("Workbook Generator")
        logger.info("=" * 80)

        # Initialize workbook
        logger.info("Creating workbook...")
        wb = Workbook()
        styles = setup_styles()

        # Create all sheets
        logger.info("Creating Dashboard...")
        create_dashboard_sheet(wb, styles)

        logger.info("Creating Screening_Process_Assessment...")
        create_screening_process_assessment_sheet(wb, styles)

        logger.info("Creating Screening_Level_Matrix...")
        create_screening_level_matrix_sheet(wb, styles)

        logger.info("Creating Personnel_Screening_Registry...")
        create_personnel_screening_registry_sheet(wb, styles)

        logger.info("Creating Screening_Compliance_Verif...")
        create_screening_compliance_verification_sheet(wb, styles)

        logger.info("Creating Continuous_Screening_Assessment...")
        create_continuous_screening_assessment_sheet(wb, styles)

        logger.info("Creating Legal_Compliance_Review...")
        create_legal_compliance_review_sheet(wb, styles)

        logger.info("Creating Gap_Analysis...")
        create_gap_analysis_sheet(wb, styles)

        logger.info("Creating Evidence_Register...")
        create_evidence_register_sheet(wb, styles)

        logger.info("Creating Approval_Sign_Off...")
        create_approval_signoff_sheet(wb, styles)

        # Set workbook properties
        wb.properties.title = "ISMS-IMP-A.5.1-2-6.1-2.S3 - Screening & Vetting Assessment"
        wb.properties.subject = "ISO/IEC 27001:2022 Control A.6.1 Assessment"
        wb.properties.creator = "[Organization] Information Security Team"
        wb.properties.keywords = "Screening, Vetting, Background Check, ISMS, ISO27001, A.6.1, FADP, GDPR"
        wb.properties.comments = "Generated via Python script generate_a5_1_2_6_1_2_s3_screening_vetting.py"

        # Generate filename with current date
        today = datetime.now().strftime("%Y%m%d")
        filename = f"ISMS-IMP-A.5.1-2-6.1-2.S3_Screening_Vetting_{today}.xlsx"

        # Save workbook
        logger.info(f"Saving workbook as: {filename}")
        wb.save(filename)

        logger.info("=" * 80)
        logger.info("Workbook generated successfully!")
        logger.info("=" * 80)
        logger.info("Next Steps:")
        logger.info("1. Open the workbook in Excel")
        logger.info("2. Complete Sheet 2 (Screening_Process_Assessment) - verify processes")
        logger.info("3. Complete Sheet 3 (Screening_Level_Matrix) - map roles to levels")
        logger.info("4. Complete Sheet 4 (Personnel_Screening_Registry) - master registry")
        logger.info("5. Complete Sheet 5 (Screening_Compliance_Verif) - per-person verification")
        logger.info("6. Complete Sheet 6 (Continuous_Screening_Assessment) - re-screening schedule")
        logger.info("7. Complete Sheet 7 (Legal_Compliance_Review) - FADP/GDPR verification")
        logger.info("8. Review Sheet 8 (Gap_Analysis) - consolidate all gaps")
        logger.info("9. Document Sheet 9 (Evidence_Register) - supporting evidence")
        logger.info("10. Review Sheet 1 (Dashboard) - auto-calculated metrics")
        logger.info("11. Obtain Sheet 10 (Approval_Sign_Off) - three-level approval")
        logger.info(f"File location: ./{filename}")

        return 0

    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
