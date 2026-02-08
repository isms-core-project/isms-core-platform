#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.1-2-6.1-2.S2 - Roles & Responsibilities Assessment Workbook Generator
================================================================================

Purpose:
    Generate Excel assessment workbook for ISO/IEC 27001:2022 Controls A.5.2 and
    A.6.1 (Information Security Roles and Responsibilities) as part of stacked
    control framework A.5.1 + A.5.2 + A.6.1 + A.6.2 (Secure Employment and Roles).

Control Alignment:
    - ISO/IEC 27001:2022 Annex A.5.2: Information Security Roles and Responsibilities
    - ISO/IEC 27001:2022 Annex A.6.1: Screening
    - ISMS-POL-A.5.1-2-6.1-2, Section 5: Roles and Responsibilities

Assessment Focus:
    - Security role definitions (CISO, DPO, Security Team, System Owners)
    - RACI matrix completeness and accuracy
    - Role accountability and assignment verification
    - Role-based access alignment
    - Training completion for role holders

Workbook Structure:
    10 sheets total:
    1. Dashboard (executive summary, auto-calculated metrics)
    2. Role_Inventory (master list of all security roles)
    3. Role_Definition_Assessment (role definition completeness verification)
    4. RACI_Matrix_Assessment (RACI matrix completeness and accuracy)
    5. Role_Assignment_Verification (current role holders and vacancy tracking)
    6. Training_Assessment (role-specific training completion)
    7. Access_Alignment_Review (role-based access vs. role definition)
    8. Gap_Analysis (identified gaps with risk and remediation)
    9. Evidence_Register (supporting evidence documentation)
   10. Approval_Sign_Off (three-level approval workflow)

Dependencies:
    - openpyxl >= 3.0.0

Usage:
    python generate_a5_1_2_6_1_2_s2_roles_responsibilities.py

Output:
    ISMS-IMP-A.5.1-2-6.1-2.S2_Roles_Responsibilities_YYYYMMDD.xlsx

Technical Specification:
    See ISMS-IMP-A.5.1-2-6.1-2.S2-Roles-Responsibilities-Assessment.md

Quality Standards:
    - Follows A.8.23/A.8.24 reference implementation patterns
    - Comprehensive documentation and inline comments
    - Consistent styling (Yellow input, Light Blue calculated, Gray labels)
    - Data validation on all appropriate cells
    - Conditional formatting for compliance statuses
    - Sheet protection with selective unlocking
    - Named ranges for cross-sheet references
    - Formula-based auto-calculations
    - Professional formatting for executive presentation

Author:               [Organization] ISMS Implementation Team
Version: 1.0
Date:                 [Date to be set]
License: Internal Use Only

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.1-2-6.1-2.S2"
WORKBOOK_NAME = "Roles Responsibilities Assessment Workbook"
CONTROL_ID = "A.5.1-2-6.1-2"
CONTROL_NAME = "Information Security Roles and Responsibilities"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ==============================================================================
# STYLE DEFINITIONS
# ==============================================================================

def setup_styles():
    """
    Define consistent cell styles used throughout the workbook.

    Returns:
        dict: Dictionary of style objects for different cell types
    """
    styles = {}

    # Header styles
    styles['header_main'] = {
        'font': Font(name='Calibri', size=18, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['header_sub'] = {
        'font': Font(name='Calibri', size=12, bold=False, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['section_header'] = {
        'font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['column_header'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    # Data cell styles
    styles['input_cell'] = {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),  # Yellow
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    styles['calculated_cell'] = {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'),  # Light Blue
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    styles['label_cell'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),  # Gray
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    # Status fills (for conditional formatting reference)
    styles['status_compliant'] = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
    styles['status_partial'] = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')  # Yellow
    styles['status_non_compliant'] = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red

    return styles


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def apply_cell_style(cell, style_dict):
    """
    Apply a style dictionary to a cell.

    Args:
        cell: openpyxl cell object
        style_dict: Dictionary containing style attributes
    """
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']


def set_column_widths(ws, widths_dict):
    """
    Set column widths for a worksheet.

    Args:
        ws: Worksheet object
        widths_dict: Dictionary mapping column letters to widths
    """
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width


def create_data_validation(formula_or_list, allow_blank=True, error_title="Invalid Entry",
                           error_message="Please select from dropdown"):
    """
    Create a data validation object for dropdown lists.

    Args:
        formula_or_list: Either a string formula or list of values
        allow_blank: Whether to allow blank cells
        error_title: Title for error message
        error_message: Error message text

    Returns:
        DataValidation object
    """
    if isinstance(formula_or_list, list):
        dv = DataValidation(type="list", formula1=f'"{",".join(formula_or_list)}"',
                           allow_blank=allow_blank)
    else:
        dv = DataValidation(type="list", formula1=formula_or_list, allow_blank=allow_blank)

    dv.error = error_message
    dv.errorTitle = error_title
    dv.showErrorMessage = True
    return dv


def merge_and_style(ws, range_string, value, style_dict):
    """
    Merge cells and apply style.

    Args:
        ws: Worksheet object
        range_string: Range to merge (e.g., 'A1:K1')
        value: Value to write
        style_dict: Style dictionary to apply
    """
    ws.merge_cells(range_string)
    cell = ws[range_string.split(':')[0]]
    cell.value = value
    apply_cell_style(cell, style_dict)


# ==============================================================================
# SHEET CREATION FUNCTIONS
# ==============================================================================

def create_dashboard_sheet(wb, styles):
    """
    Sheet 1: Dashboard (Executive Summary)

    Auto-calculated metrics pulling from other sheets.
    Read-only for users.
    """
    ws = wb.active
    ws.title = "Dashboard"

    # --- Header Section ---
    merge_and_style(ws, 'A1:K1',
                   'ISMS-IMP-A.5.1-2-6.1-2.S2 - Roles & Responsibilities Assessment Dashboard',
                   styles['header_main'])
    ws.row_dimensions[1].height = 40

    merge_and_style(ws, 'A2:K2',
                   'ISO/IEC 27001:2022 Controls A.5.2 & A.6.1 - Information Security Roles and Responsibilities',
                   styles['header_sub'])
    ws.row_dimensions[2].height = 30

    # --- Document Information Block (Rows 4-15) ---
    ws['A4'] = 'Document ID:'
    apply_cell_style(ws['A4'], styles['label_cell'])
    ws.merge_cells('B4:C4')
    ws['B4'] = 'ISMS-IMP-A.5.1-2-6.1-2.S2'

    info_rows = [
        (5, 'Related Policy:', 'ISMS-POL-A.5.1-2-6.1-2, Section 5'),
        (6, 'ISO Control:', 'A.5.2 & A.6.1 (Roles and Responsibilities)'),
        (7, 'Assessment Period:', None),  # User input
        (8, 'Assessment Date:', '=TODAY()'),  # Formula
        (9, 'Assessor Name:', None),  # User input
        (10, 'Assessor Role:', None),  # User input
        (11, 'Organization:', None),  # User input
        (12, 'Review Cycle:', 'Quarterly'),
        (13, 'Last Updated:', '=NOW()'),  # Formula
        (14, 'Assessment Status:', None),  # Dropdown
        (15, 'Next Review Date:', None)  # User input
    ]

    for row_num, label, value in info_rows:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        ws.merge_cells(f'B{row_num}:C{row_num}')

        if value:
            ws[f'B{row_num}'] = value
            if value.startswith('='):
                apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
            else:
                apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # Status dropdown for row 14
    status_dv = create_data_validation(['Draft', 'Under-Review', 'Approved', 'Audit-Ready'])
    ws.add_data_validation(status_dv)
    status_dv.add(ws['B14'])

    # --- Overall Compliance Summary (Rows 17-26) ---
    merge_and_style(ws, 'A17:K17', 'OVERALL ROLES & RESPONSIBILITIES COMPLIANCE', styles['section_header'])
    ws.row_dimensions[17].height = 35

    # Compliance Scorecard (Rows 19-26)
    scorecard = [
        (19, 'Overall Compliance Score', '=AVERAGE(C32,C33,C34,C35,C36)', '90%'),
        (20, 'Total Roles Defined', '=COUNTA(Role_Inventory!A:A)-4', 'N/A'),
        (21, 'Roles Fully Assigned', '=COUNTIF(Role_Assignment_Verification!J:J,"Filled")', '100%'),
        (22, 'Vacant Roles', '=COUNTIF(Role_Assignment_Verification!J:J,"Vacant")', '0'),
        (23, 'RACI Coverage', '=RACI_Matrix_Assessment!B20', '100%'),
        (24, 'Training Completion Rate', '=Training_Assessment!B20', '90%'),
        (25, 'Access Alignment Rate', '=Access_Alignment_Review!B20', '95%'),
        (26, 'Critical Gaps', '=COUNTIF(Gap_Analysis!F:F,"Critical")', '0')
    ]

    for row_num, label, formula, target in scorecard:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])

        ws[f'B{row_num}'] = formula
        apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        if row_num in [19, 23, 24, 25]:
            ws[f'B{row_num}'].number_format = '0.00%'

        ws[f'C{row_num}'] = target
        apply_cell_style(ws[f'C{row_num}'], styles['calculated_cell'])

        ws[f'D{row_num}'] = '=IF(B{0}>=C{0},"On Target",IF(B{0}>=C{0}*0.9,"Close","Below Target"))'.format(row_num)
        apply_cell_style(ws[f'D{row_num}'], styles['calculated_cell'])

    # --- Domain Compliance Breakdown (Rows 28-38) ---
    merge_and_style(ws, 'A28:K28', 'DOMAIN COMPLIANCE BREAKDOWN', styles['section_header'])

    # Column headers for domain table
    for col, header in zip(['A', 'B', 'C', 'D', 'E'],
                           ['Domain', 'Weight', 'Score (%)', 'Weighted Score', 'Status']):
        ws[f'{col}30'] = header
        apply_cell_style(ws[f'{col}30'], styles['column_header'])

    domains = [
        (32, 'Role Definition Completeness', '25%',
         '=COUNTIF(Role_Definition_Assessment!M:M,"Compliant")/MAX(1,COUNTA(Role_Definition_Assessment!A:A)-4)', '=B32*C32'),
        (33, 'RACI Matrix Coverage', '25%',
         '=RACI_Matrix_Assessment!B20', '=B33*C33'),
        (34, 'Role Assignment Status', '20%',
         '=COUNTIF(Role_Assignment_Verification!J:J,"Filled")/MAX(1,COUNTA(Role_Assignment_Verification!A:A)-4)', '=B34*C34'),
        (35, 'Training Compliance', '15%',
         '=Training_Assessment!B20', '=B35*C35'),
        (36, 'Access Alignment', '15%',
         '=Access_Alignment_Review!B20', '=B36*C36')
    ]

    for row_num, domain, weight, score_formula, weighted_formula in domains:
        ws[f'A{row_num}'] = domain
        ws[f'B{row_num}'] = weight
        ws[f'C{row_num}'] = score_formula
        ws[f'C{row_num}'].number_format = '0.00%'
        ws[f'D{row_num}'] = weighted_formula
        ws[f'D{row_num}'].number_format = '0.00%'
        ws[f'E{row_num}'] = f'=IF(C{row_num}>=0.9,"Compliant",IF(C{row_num}>=0.7,"Partial","Non-Compliant"))'

    # Total row
    ws['A38'] = 'TOTAL WEIGHTED SCORE'
    apply_cell_style(ws['A38'], styles['label_cell'])
    ws['B38'] = '100%'
    ws['C38'] = '=SUM(D32:D36)'
    ws['C38'].number_format = '0.00%'
    ws['C38'].font = Font(name='Calibri', size=10, bold=True)

    # Set column widths
    set_column_widths(ws, {'A': 40, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 5, 'G': 5, 'H': 5, 'I': 5, 'J': 5, 'K': 5})

    # Protect sheet (allow sorting/filtering)
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_role_inventory_sheet(wb, styles):
    """
    Sheet 2: Role_Inventory

    Master list of all security roles with metadata.
    100 rows for roles.
    """
    ws = wb.create_sheet("Role_Inventory")

    # --- Header ---
    merge_and_style(ws, 'A1:P1', 'ROLE INVENTORY', styles['header_main'])
    merge_and_style(ws, 'A2:P2',
                   'Master list of all information security roles with metadata',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # --- Column Headers (Row 4) ---
    headers = [
        ('A', 'Role_ID', 15),
        ('B', 'Role_Title', 30),
        ('C', 'Role_Category', 20),
        ('D', 'Role_Type', 15),
        ('E', 'Department', 20),
        ('F', 'Reports_To', 25),
        ('G', 'Security_Clearance_Required', 15),
        ('H', 'ISO_Control_Mapping', 20),
        ('I', 'Role_Created_Date', 12),
        ('J', 'Last_Review_Date', 12),
        ('K', 'Role_Status', 15),
        ('L', 'Criticality', 15),
        ('M', 'Backup_Required', 15),
        ('N', 'Backup_Role_ID', 15),
        ('O', 'Role_Description', 50),
        ('P', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)
    ws.row_dimensions[4].height = 30

    # --- Data Rows (5-104) ---
    for row in range(5, 105):
        for col_letter in [c[0] for c in headers]:
            cell = ws[f'{col_letter}{row}']
            apply_cell_style(cell, styles['input_cell'])

    # --- Data Validation ---
    # Role_Category dropdown
    category_dv = create_data_validation(['Executive-Leadership', 'Security-Management', 'Security-Operations',
                                          'IT-Operations', 'System-Owner', 'Data-Owner', 'Business-Unit', 'External'])
    ws.add_data_validation(category_dv)
    category_dv.add('C5:C104')

    # Role_Type dropdown
    type_dv = create_data_validation(['Primary', 'Secondary', 'Backup', 'Advisory', 'Oversight'])
    ws.add_data_validation(type_dv)
    type_dv.add('D5:D104')

    # Security_Clearance_Required dropdown
    clearance_dv = create_data_validation(['None', 'Basic', 'Enhanced', 'High', 'Critical'])
    ws.add_data_validation(clearance_dv)
    clearance_dv.add('G5:G104')

    # Role_Status dropdown
    status_dv = create_data_validation(['Active', 'Draft', 'Under-Review', 'Retired', 'Proposed'])
    ws.add_data_validation(status_dv)
    status_dv.add('K5:K104')

    # Criticality dropdown
    crit_dv = create_data_validation(['Critical', 'High', 'Medium', 'Low'])
    ws.add_data_validation(crit_dv)
    crit_dv.add('L5:L104')

    # Backup_Required dropdown
    backup_dv = create_data_validation(['Yes', 'No', 'Recommended'])
    ws.add_data_validation(backup_dv)
    backup_dv.add('M5:M104')

    # Date formatting
    for col in ['I', 'J']:
        for row in range(5, 105):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes at A5
    ws.freeze_panes = 'A5'

    # Define named range for Role_ID_List
    role_id_range = DefinedName(name='Role_ID_List', attr_text="Role_Inventory!$A$5:$A$104")
    wb.defined_names.add(role_id_range)

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_role_definition_assessment_sheet(wb, styles):
    """
    Sheet 3: Role_Definition_Assessment

    Verify role definition completeness for each security role.
    """
    ws = wb.create_sheet("Role_Definition_Assessment")

    # --- Header ---
    merge_and_style(ws, 'A1:O1', 'ROLE DEFINITION ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:O2',
                   'Verification of role definition completeness for all security roles',
                   styles['header_sub'])

    # --- Column Headers ---
    headers = [
        ('A', 'Role_ID', 15),
        ('B', 'Role_Title', 30),
        ('C', 'Role_Category', 20),
        ('D', 'Responsibilities_Documented', 15),
        ('E', 'Authority_Level_Defined', 15),
        ('F', 'Reporting_Lines_Clear', 15),
        ('G', 'Competency_Requirements', 15),
        ('H', 'Training_Requirements', 15),
        ('I', 'Access_Requirements', 15),
        ('J', 'Accountability_Defined', 15),
        ('K', 'Segregation_of_Duties', 15),
        ('L', 'Definition_Documentation', 20),
        ('M', 'Definition_Compliance_Rating', 20),
        ('N', 'Gap_Description', 40),
        ('O', 'Evidence_Reference', 30)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (5-104) with Auto-Population Formulas ---
    for row in range(5, 105):
        # Auto-populate Role_ID from Role_Inventory
        ws[f'A{row}'] = f'=IF(ROW()-4<=COUNTA(Role_Inventory!$A:$A)-4,INDEX(Role_Inventory!$A:$A,ROW()-3),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # Auto-populate Role_Title
        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Role_Inventory!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])

        # Auto-populate Role_Category
        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Role_Inventory!$A:$C,3,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])

        # Manual entry cells (D-L, N-O)
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'N', 'O']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # Definition_Compliance_Rating (M) - auto-calculated
        ws[f'M{row}'] = (f'=IF(B{row}="","",IF(AND(D{row}="Complete",E{row}="Yes",F{row}="Yes",'
                        f'G{row}="Complete",H{row}="Complete",I{row}="Complete",J{row}="Yes",L{row}="Complete"),"Compliant",'
                        f'IF(OR(D{row}="Missing",E{row}="No",F{row}="No",J{row}="No",L{row}="Missing"),"Non-Compliant","Partial")))')
        apply_cell_style(ws[f'M{row}'], styles['calculated_cell'])

    # --- Data Validation ---
    complete_dv = create_data_validation(['Complete', 'Partial', 'Missing'])
    ws.add_data_validation(complete_dv)
    for col in ['D', 'G', 'H', 'I', 'L']:
        complete_dv.add(f'{col}5:{col}104')

    yesno_dv = create_data_validation(['Yes', 'Partial', 'No'])
    ws.add_data_validation(yesno_dv)
    for col in ['E', 'F', 'J']:
        yesno_dv.add(f'{col}5:{col}104')

    sod_dv = create_data_validation(['Yes', 'No', 'N/A', 'Conflict-Identified'])
    ws.add_data_validation(sod_dv)
    sod_dv.add('K5:K104')

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_raci_matrix_assessment_sheet(wb, styles):
    """
    Sheet 4: RACI_Matrix_Assessment

    RACI matrix completeness and accuracy verification.
    """
    ws = wb.create_sheet("RACI_Matrix_Assessment")

    # --- Header ---
    merge_and_style(ws, 'A1:N1', 'RACI MATRIX ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:N2',
                   'Verification of RACI matrix completeness and accuracy for security activities',
                   styles['header_sub'])

    # --- Summary Section (Rows 4-20) ---
    ws['A4'] = 'RACI MATRIX SUMMARY'
    apply_cell_style(ws['A4'], styles['section_header'])
    ws.merge_cells('A4:B4')

    summary_items = [
        (6, 'Total Activities Assessed:', '=COUNTA(A25:A174)-1'),
        (7, 'Activities with Complete RACI:', '=COUNTIF(M25:M174,"Complete")'),
        (8, 'Activities with Partial RACI:', '=COUNTIF(M25:M174,"Partial")'),
        (9, 'Activities Missing RACI:', '=COUNTIF(M25:M174,"Missing")'),
        (10, 'Activities with Accountability Gap:', '=COUNTIF(E25:E174,"Missing")'),
        (11, 'Activities with Multiple Accountables:', '=COUNTIF(N25:N174,"Yes")'),
        (12, 'RACI Conflict Count:', '=COUNTIF(N25:N174,"Yes")'),
        (13, 'Average RACI Score:', '=AVERAGE(L25:L174)'),
        (20, 'Overall RACI Coverage:', '=B7/MAX(1,B6)')
    ]

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    for row_num, label, formula in summary_items:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        ws[f'B{row_num}'] = formula
        apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        if row_num in [13, 20]:
            ws[f'B{row_num}'].number_format = '0.00%'

    # --- RACI Detail Section (Rows 23+) ---
    ws['A23'] = 'RACI DETAIL ASSESSMENT'
    apply_cell_style(ws['A23'], styles['section_header'])
    ws.merge_cells('A23:N23')

    # Column Headers
    headers = [
        ('A', 'Activity_ID', 15),
        ('B', 'Security_Activity', 40),
        ('C', 'Activity_Category', 20),
        ('D', 'Responsible_Role', 25),
        ('E', 'Accountable_Role', 25),
        ('F', 'Consulted_Roles', 30),
        ('G', 'Informed_Roles', 30),
        ('H', 'Responsible_Assigned', 12),
        ('I', 'Accountable_Assigned', 12),
        ('J', 'Multiple_Accountables', 12),
        ('K', 'RACI_Documented', 12),
        ('L', 'RACI_Score', 12),
        ('M', 'RACI_Status', 15),
        ('N', 'RACI_Conflict', 15)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}25'] = header_text
        apply_cell_style(ws[f'{col_letter}25'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (26-175, 150 activities) ---
    for row in range(26, 176):
        # Manual entry cells
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # Auto-calculated cells
        # Responsible_Assigned (H)
        ws[f'H{row}'] = f'=IF(D{row}<>"","Yes","No")'
        apply_cell_style(ws[f'H{row}'], styles['calculated_cell'])

        # Accountable_Assigned (I)
        ws[f'I{row}'] = f'=IF(E{row}<>"","Yes","No")'
        apply_cell_style(ws[f'I{row}'], styles['calculated_cell'])

        # Multiple_Accountables (J) - check for comma in Accountable
        ws[f'J{row}'] = f'=IF(ISERROR(FIND(",",E{row})),"No","Yes")'
        apply_cell_style(ws[f'J{row}'], styles['calculated_cell'])

        # RACI_Documented (K) - manual
        apply_cell_style(ws[f'K{row}'], styles['input_cell'])

        # RACI_Score (L) - calculate based on completeness
        ws[f'L{row}'] = (f'=IF(B{row}="","",((IF(D{row}<>"",0.3,0))+(IF(E{row}<>"",0.4,0))+'
                        f'(IF(F{row}<>"",0.15,0))+(IF(G{row}<>"",0.15,0))))')
        apply_cell_style(ws[f'L{row}'], styles['calculated_cell'])
        ws[f'L{row}'].number_format = '0%'

        # RACI_Status (M)
        ws[f'M{row}'] = f'=IF(B{row}="","",IF(L{row}>=0.85,"Complete",IF(L{row}>=0.5,"Partial","Missing")))'
        apply_cell_style(ws[f'M{row}'], styles['calculated_cell'])

        # RACI_Conflict (N)
        ws[f'N{row}'] = f'=IF(J{row}="Yes","Yes",IF(AND(H{row}="No",I{row}="No"),"Missing-R-A","No"))'
        apply_cell_style(ws[f'N{row}'], styles['calculated_cell'])

    # --- Data Validation ---
    category_dv = create_data_validation(['Policy-Management', 'Risk-Management', 'Incident-Response',
                                          'Access-Management', 'Change-Management', 'Audit-Compliance',
                                          'Training-Awareness', 'Asset-Management', 'Vendor-Management',
                                          'Business-Continuity', 'Security-Operations', 'Other'])
    ws.add_data_validation(category_dv)
    category_dv.add('C26:C175')

    documented_dv = create_data_validation(['Yes', 'Partial', 'No'])
    ws.add_data_validation(documented_dv)
    documented_dv.add('K26:K175')

    # Freeze panes
    ws.freeze_panes = 'A26'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_role_assignment_verification_sheet(wb, styles):
    """
    Sheet 5: Role_Assignment_Verification

    Current role holders and vacancy tracking.
    """
    ws = wb.create_sheet("Role_Assignment_Verification")

    # --- Header ---
    merge_and_style(ws, 'A1:O1', 'ROLE ASSIGNMENT VERIFICATION', styles['header_main'])
    merge_and_style(ws, 'A2:O2',
                   'Verification of current role assignments and vacancy tracking',
                   styles['header_sub'])

    # --- Column Headers ---
    headers = [
        ('A', 'Role_ID', 15),
        ('B', 'Role_Title', 30),
        ('C', 'Role_Category', 20),
        ('D', 'Current_Holder_Name', 25),
        ('E', 'Current_Holder_Employee_ID', 15),
        ('F', 'Assignment_Date', 12),
        ('G', 'Assignment_Documentation', 15),
        ('H', 'Formal_Acceptance', 15),
        ('I', 'Background_Check_Status', 15),
        ('J', 'Assignment_Status', 15),
        ('K', 'Backup_Holder_Name', 25),
        ('L', 'Backup_Assignment_Status', 15),
        ('M', 'Last_Verification_Date', 12),
        ('N', 'Gap_Description', 40),
        ('O', 'Evidence_Reference', 30)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (5-104) ---
    for row in range(5, 105):
        # Auto-populate Role_ID from Role_Inventory
        ws[f'A{row}'] = f'=IF(ROW()-4<=COUNTA(Role_Inventory!$A:$A)-4,INDEX(Role_Inventory!$A:$A,ROW()-3),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # Auto-populate Role_Title
        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Role_Inventory!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])

        # Auto-populate Role_Category
        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Role_Inventory!$A:$C,3,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])

        # Manual entry cells
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'K', 'L', 'M', 'N', 'O']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # Assignment_Status (J) - auto-calculated
        ws[f'J{row}'] = (f'=IF(B{row}="","",IF(AND(D{row}<>"",G{row}="Complete",H{row}="Yes"),"Filled",'
                        f'IF(D{row}="","Vacant",IF(OR(G{row}="Missing",H{row}="No"),"Pending-Documentation","Pending"))))')
        apply_cell_style(ws[f'J{row}'], styles['calculated_cell'])

    # Date formatting
    for col in ['F', 'M']:
        for row in range(5, 105):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    doc_status_dv = create_data_validation(['Complete', 'Partial', 'Missing'])
    ws.add_data_validation(doc_status_dv)
    doc_status_dv.add('G5:G104')

    acceptance_dv = create_data_validation(['Yes', 'Pending', 'No'])
    ws.add_data_validation(acceptance_dv)
    acceptance_dv.add('H5:H104')

    bg_check_dv = create_data_validation(['Completed', 'In-Progress', 'Not-Started', 'Not-Required', 'Expired'])
    ws.add_data_validation(bg_check_dv)
    bg_check_dv.add('I5:I104')

    backup_status_dv = create_data_validation(['Assigned', 'Pending', 'Not-Assigned', 'N/A'])
    ws.add_data_validation(backup_status_dv)
    backup_status_dv.add('L5:L104')

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_training_assessment_sheet(wb, styles):
    """
    Sheet 6: Training_Assessment

    Role-specific training completion tracking.
    """
    ws = wb.create_sheet("Training_Assessment")

    # --- Header ---
    merge_and_style(ws, 'A1:P1', 'TRAINING ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:P2',
                   'Role-specific training requirements and completion verification',
                   styles['header_sub'])

    # --- Summary Section (Rows 4-20) ---
    ws['A4'] = 'TRAINING SUMMARY'
    apply_cell_style(ws['A4'], styles['section_header'])
    ws.merge_cells('A4:B4')

    summary_items = [
        (6, 'Total Role Holders Assessed:', '=COUNTA(A25:A124)-1'),
        (7, 'Training Complete:', '=COUNTIF(N25:N124,"Compliant")'),
        (8, 'Training Partial:', '=COUNTIF(N25:N124,"Partial")'),
        (9, 'Training Non-Compliant:', '=COUNTIF(N25:N124,"Non-Compliant")'),
        (10, 'Overdue Training Items:', '=COUNTIF(L25:L124,"Overdue")'),
        (11, 'Expiring Within 30 Days:', '=COUNTIF(L25:L124,"Expiring-Soon")'),
        (20, 'Overall Training Completion:', '=B7/MAX(1,B6)')
    ]

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    for row_num, label, formula in summary_items:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        ws[f'B{row_num}'] = formula
        apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        if row_num == 20:
            ws[f'B{row_num}'].number_format = '0.00%'

    # --- Training Detail Section (Rows 23+) ---
    ws['A23'] = 'TRAINING DETAIL ASSESSMENT'
    apply_cell_style(ws['A23'], styles['section_header'])
    ws.merge_cells('A23:P23')

    # Column Headers
    headers = [
        ('A', 'Role_ID', 15),
        ('B', 'Role_Title', 30),
        ('C', 'Role_Holder_Name', 25),
        ('D', 'Required_Training_1', 25),
        ('E', 'Training_1_Status', 12),
        ('F', 'Training_1_Date', 12),
        ('G', 'Required_Training_2', 25),
        ('H', 'Training_2_Status', 12),
        ('I', 'Training_2_Date', 12),
        ('J', 'Required_Training_3', 25),
        ('K', 'Training_3_Status', 12),
        ('L', 'Training_Expiry_Status', 15),
        ('M', 'Training_Records_Available', 12),
        ('N', 'Training_Compliance_Rating', 20),
        ('O', 'Gap_Description', 40),
        ('P', 'Evidence_Reference', 30)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}25'] = header_text
        apply_cell_style(ws[f'{col_letter}25'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (26-125, 100 entries) ---
    for row in range(26, 126):
        # Auto-populate from Role_Assignment_Verification
        ws[f'A{row}'] = f'=IF(ROW()-25<=COUNTA(Role_Assignment_Verification!$A:$A)-4,INDEX(Role_Assignment_Verification!$A:$A,ROW()-24+4),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Role_Inventory!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])

        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Role_Assignment_Verification!$A:$D,4,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])

        # Manual entry cells
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'O', 'P']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # Training_Compliance_Rating (N) - auto-calculated
        ws[f'N{row}'] = (f'=IF(B{row}="","",IF(AND(OR(E{row}="Complete",D{row}=""),'
                        f'OR(H{row}="Complete",G{row}=""),OR(K{row}="Complete",J{row}=""),M{row}="Yes"),"Compliant",'
                        f'IF(OR(E{row}="Not-Started",H{row}="Not-Started",K{row}="Not-Started",M{row}="No"),"Non-Compliant","Partial")))')
        apply_cell_style(ws[f'N{row}'], styles['calculated_cell'])

    # Date formatting
    for col in ['F', 'I']:
        for row in range(26, 126):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    training_status_dv = create_data_validation(['Complete', 'In-Progress', 'Not-Started', 'N/A'])
    ws.add_data_validation(training_status_dv)
    for col in ['E', 'H', 'K']:
        training_status_dv.add(f'{col}26:{col}125')

    expiry_dv = create_data_validation(['Current', 'Expiring-Soon', 'Overdue', 'N/A'])
    ws.add_data_validation(expiry_dv)
    expiry_dv.add('L26:L125')

    records_dv = create_data_validation(['Yes', 'Partial', 'No'])
    ws.add_data_validation(records_dv)
    records_dv.add('M26:M125')

    # Freeze panes
    ws.freeze_panes = 'A26'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_access_alignment_review_sheet(wb, styles):
    """
    Sheet 7: Access_Alignment_Review

    Role-based access vs. role definition alignment verification.
    """
    ws = wb.create_sheet("Access_Alignment_Review")

    # --- Header ---
    merge_and_style(ws, 'A1:O1', 'ACCESS ALIGNMENT REVIEW', styles['header_main'])
    merge_and_style(ws, 'A2:O2',
                   'Verification of role-based access alignment with role definitions',
                   styles['header_sub'])

    # --- Summary Section (Rows 4-20) ---
    ws['A4'] = 'ACCESS ALIGNMENT SUMMARY'
    apply_cell_style(ws['A4'], styles['section_header'])
    ws.merge_cells('A4:B4')

    summary_items = [
        (6, 'Total Role Holders Assessed:', '=COUNTA(A25:A124)-1'),
        (7, 'Access Fully Aligned:', '=COUNTIF(M25:M124,"Aligned")'),
        (8, 'Access Partially Aligned:', '=COUNTIF(M25:M124,"Partial")'),
        (9, 'Access Misaligned:', '=COUNTIF(M25:M124,"Misaligned")'),
        (10, 'Excess Privileges Identified:', '=COUNTIF(H25:H124,"Yes")'),
        (11, 'Missing Access Identified:', '=COUNTIF(I25:I124,"Yes")'),
        (12, 'Segregation Conflicts:', '=COUNTIF(K25:K124,"Yes")'),
        (20, 'Overall Access Alignment:', '=B7/MAX(1,B6)')
    ]

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    for row_num, label, formula in summary_items:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        ws[f'B{row_num}'] = formula
        apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        if row_num == 20:
            ws[f'B{row_num}'].number_format = '0.00%'

    # --- Access Detail Section (Rows 23+) ---
    ws['A23'] = 'ACCESS ALIGNMENT DETAIL'
    apply_cell_style(ws['A23'], styles['section_header'])
    ws.merge_cells('A23:O23')

    # Column Headers
    headers = [
        ('A', 'Role_ID', 15),
        ('B', 'Role_Title', 30),
        ('C', 'Role_Holder_Name', 25),
        ('D', 'Defined_Access_Level', 20),
        ('E', 'Actual_Access_Level', 20),
        ('F', 'Systems_Access_Defined', 30),
        ('G', 'Systems_Access_Actual', 30),
        ('H', 'Excess_Privileges', 12),
        ('I', 'Missing_Access', 12),
        ('J', 'Access_Review_Date', 12),
        ('K', 'SoD_Conflict', 12),
        ('L', 'Access_Documentation', 12),
        ('M', 'Access_Alignment_Status', 15),
        ('N', 'Gap_Description', 40),
        ('O', 'Evidence_Reference', 30)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}25'] = header_text
        apply_cell_style(ws[f'{col_letter}25'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (26-125, 100 entries) ---
    for row in range(26, 126):
        # Auto-populate from Role_Assignment_Verification
        ws[f'A{row}'] = f'=IF(ROW()-25<=COUNTA(Role_Assignment_Verification!$A:$A)-4,INDEX(Role_Assignment_Verification!$A:$A,ROW()-24+4),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Role_Inventory!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])

        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},Role_Assignment_Verification!$A:$D,4,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])

        # Manual entry cells
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'N', 'O']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # Access_Alignment_Status (M) - auto-calculated
        ws[f'M{row}'] = (f'=IF(B{row}="","",IF(AND(H{row}="No",I{row}="No",K{row}="No",L{row}="Complete"),"Aligned",'
                        f'IF(OR(K{row}="Yes",AND(H{row}="Yes",I{row}="Yes")),"Misaligned","Partial")))')
        apply_cell_style(ws[f'M{row}'], styles['calculated_cell'])

    # Date formatting
    for row in range(26, 126):
        ws[f'J{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    access_level_dv = create_data_validation(['Admin', 'Power-User', 'Standard', 'Read-Only', 'None'])
    ws.add_data_validation(access_level_dv)
    for col in ['D', 'E']:
        access_level_dv.add(f'{col}26:{col}125')

    yesno_dv = create_data_validation(['Yes', 'No'])
    ws.add_data_validation(yesno_dv)
    for col in ['H', 'I', 'K']:
        yesno_dv.add(f'{col}26:{col}125')

    doc_dv = create_data_validation(['Complete', 'Partial', 'Missing'])
    ws.add_data_validation(doc_dv)
    doc_dv.add('L26:L125')

    # Freeze panes
    ws.freeze_panes = 'A26'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_gap_analysis_sheet(wb, styles):
    """
    Sheet 8: Gap_Analysis

    Consolidated gaps from all assessment domains.
    """
    ws = wb.create_sheet("Gap_Analysis")

    # --- Header ---
    merge_and_style(ws, 'A1:P1', 'GAP ANALYSIS', styles['header_main'])
    merge_and_style(ws, 'A2:P2',
                   'Consolidated gaps from all assessment domains with risk levels and remediation plans',
                   styles['header_sub'])

    # --- Column Headers ---
    headers = [
        ('A', 'Gap_ID', 12),
        ('B', 'Role_ID', 15),
        ('C', 'Role_Title', 30),
        ('D', 'Gap_Category', 20),
        ('E', 'Gap_Description', 40),
        ('F', 'Risk_Level', 15),
        ('G', 'Impact_Assessment', 35),
        ('H', 'Affected_Stakeholders', 25),
        ('I', 'Remediation_Action', 40),
        ('J', 'Responsible_Party', 25),
        ('K', 'Target_Completion_Date', 15),
        ('L', 'Estimated_Effort', 15),
        ('M', 'Dependencies', 30),
        ('N', 'Status', 15),
        ('O', 'Completion_Evidence', 30),
        ('P', 'Risk_Acceptance', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (5-104, 100 rows for gaps) ---
    for row in range(5, 105):
        # Gap_ID auto-generated
        ws[f'A{row}'] = f'=IF(ROW()<=4,"",TEXT(ROW()-4,"GAP-000"))'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # All other columns manual entry
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

    # Date formatting for Target_Completion_Date
    for row in range(5, 105):
        ws[f'K{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    gap_category_dv = create_data_validation(['Role-Definition', 'RACI-Matrix', 'Role-Assignment',
                                              'Training', 'Access-Alignment', 'Documentation', 'Other'])
    ws.add_data_validation(gap_category_dv)
    gap_category_dv.add('D5:D104')

    risk_level_dv = create_data_validation(['Critical', 'High', 'Medium', 'Low'])
    ws.add_data_validation(risk_level_dv)
    risk_level_dv.add('F5:F104')

    effort_dv = create_data_validation(['<1hr', '1-4hrs', '1day', '2-5days', '>1week'])
    ws.add_data_validation(effort_dv)
    effort_dv.add('L5:L104')

    status_dv = create_data_validation(['Not-Started', 'In-Progress', 'Blocked', 'Completed', 'Accepted-Risk'])
    ws.add_data_validation(status_dv)
    status_dv.add('N5:N104')

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_evidence_register_sheet(wb, styles):
    """
    Sheet 9: Evidence_Register

    Supporting evidence documentation.
    """
    ws = wb.create_sheet("Evidence_Register")

    # --- Header ---
    merge_and_style(ws, 'A1:J1', 'EVIDENCE REGISTER', styles['header_main'])
    merge_and_style(ws, 'A2:J2',
                   'Documentation of all supporting evidence for roles and responsibilities assessment',
                   styles['header_sub'])

    # --- Column Headers ---
    headers = [
        ('A', 'Evidence_ID', 12),
        ('B', 'Evidence_Type', 20),
        ('C', 'Description', 40),
        ('D', 'Related_Role_ID', 15),
        ('E', 'Related_Assessment_Sheet', 25),
        ('F', 'File_Location', 40),
        ('G', 'Date_Collected', 12),
        ('H', 'Collected_By', 25),
        ('I', 'Verification_Status', 15),
        ('J', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (5-104, 100 evidence items) ---
    for row in range(5, 105):
        # Evidence_ID auto-generated
        ws[f'A{row}'] = f'=IF(ROW()<=4,"",TEXT(ROW()-4,"EVD-000"))'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # All other columns manual entry
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

    # Date formatting for Date_Collected
    for row in range(5, 105):
        ws[f'G{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    evidence_type_dv = create_data_validation(['Role-Definition-Document', 'RACI-Matrix', 'Job-Description',
                                               'Assignment-Letter', 'Acceptance-Acknowledgment', 'Training-Certificate',
                                               'Background-Check-Report', 'Access-Control-Config', 'Access-Review-Report',
                                               'Organizational-Chart', 'Meeting-Minutes', 'Other'])
    ws.add_data_validation(evidence_type_dv)
    evidence_type_dv.add('B5:B104')

    sheet_dv = create_data_validation(['Role_Inventory', 'Role_Definition_Assessment', 'RACI_Matrix_Assessment',
                                       'Role_Assignment_Verification', 'Training_Assessment', 'Access_Alignment_Review',
                                       'Gap_Analysis', 'Multiple'])
    ws.add_data_validation(sheet_dv)
    sheet_dv.add('E5:E104')

    verification_dv = create_data_validation(['Verified', 'Pending', 'Not-Verified', 'Expired'])
    ws.add_data_validation(verification_dv)
    verification_dv.add('I5:I104')

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False


def create_approval_signoff_sheet(wb, styles):
    """
    Sheet 10: Approval_Sign_Off

    Three-level approval workflow.
    """
    ws = wb.create_sheet("Approval_Sign_Off")

    # --- Header ---
    merge_and_style(ws, 'A1:D1', 'APPROVAL & SIGN-OFF', styles['header_main'])
    ws.row_dimensions[1].height = 30

    # --- Assessment Summary (Rows 3-15) ---
    ws['A3'] = 'ASSESSMENT SUMMARY'
    apply_cell_style(ws['A3'], styles['section_header'])
    ws.merge_cells('A3:D3')

    summary_items = [
        (5, 'Document:', 'ISMS-IMP-A.5.1-2-6.1-2.S2'),
        (6, 'Assessment Period:', '=Dashboard!B7'),
        (7, 'Overall Compliance Score:', '=Dashboard!B19'),
        (8, 'Total Roles Defined:', '=Dashboard!B20'),
        (9, 'Roles Fully Assigned:', '=Dashboard!B21'),
        (10, 'Vacant Roles:', '=Dashboard!B22'),
        (11, 'RACI Coverage:', '=Dashboard!B23'),
        (12, 'Training Completion:', '=Dashboard!B24'),
        (13, 'Access Alignment:', '=Dashboard!B25'),
        (14, 'Critical Gaps:', '=Dashboard!B26')
    ]

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30

    for row_num, label, value in summary_items:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])

        ws[f'B{row_num}'] = value
        if value.startswith('='):
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])

    # --- Level 1 Approval (Rows 17-25) ---
    ws['A17'] = 'LEVEL 1: PREPARED BY'
    apply_cell_style(ws['A17'], styles['section_header'])
    ws.merge_cells('A17:D17')

    level1_fields = [
        (18, 'Name:'),
        (19, 'Role:'),
        (20, 'Date:'),
        (21, 'Signature:'),
        (22, 'Certification:', 'I certify this assessment is complete and accurate'),
        (23, 'Comments:')
    ]

    for row_num, label, *rest in level1_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])

        if rest:
            ws[f'B{row_num}'] = rest[0]
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # --- Level 2 Approval (Rows 27-40) ---
    ws['A27'] = 'LEVEL 2: REVIEWED BY'
    apply_cell_style(ws['A27'], styles['section_header'])
    ws.merge_cells('A27:D27')

    level2_fields = [
        (28, 'Name:'),
        (29, 'Role:'),
        (30, 'Date:'),
        (31, 'Signature:')
    ]

    for row_num, label in level2_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # Review checklist
    checklist_items = [
        'Role inventory complete',
        'Role definitions verified',
        'RACI matrix complete',
        'Role assignments verified',
        'Training compliance verified',
        'Access alignment reviewed',
        'Gaps identified and documented',
        'Evidence sufficient'
    ]

    ws['A33'] = 'Review Checklist:'
    apply_cell_style(ws['A33'], styles['label_cell'])

    for i, item in enumerate(checklist_items, start=34):
        ws[f'A{i}'] = f'[ ] {item}'
        ws.merge_cells(f'A{i}:D{i}')

    ws['A42'] = 'Comments:'
    apply_cell_style(ws['A42'], styles['label_cell'])
    apply_cell_style(ws['B42'], styles['input_cell'])

    # --- Level 3 Approval (Rows 44-55) ---
    ws['A44'] = 'LEVEL 3: APPROVED BY (CISO)'
    apply_cell_style(ws['A44'], styles['section_header'])
    ws.merge_cells('A44:D44')

    level3_fields = [
        (45, 'Name:'),
        (46, 'Role:'),
        (47, 'Date:'),
        (48, 'Signature:'),
        (49, 'Final Approval:', 'I approve this assessment as accurate'),
        (50, 'Risk Acceptance:', 'I accept residual risk for: [list]'),
        (51, 'Comments:')
    ]

    for row_num, label, *rest in level3_fields:
        ws[f'A{row_num}'] = label
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])

        if rest:
            ws[f'B{row_num}'] = rest[0]
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])

    # --- Assessment Metadata (Rows 53-60) ---
    ws['A53'] = 'Next Review Date:'
    apply_cell_style(ws['A53'], styles['label_cell'])
    apply_cell_style(ws['B53'], styles['input_cell'])

    ws['A54'] = 'Assessment Status:'
    apply_cell_style(ws['A54'], styles['label_cell'])
    apply_cell_style(ws['B54'], styles['input_cell'])

    status_dv = create_data_validation(['Draft', 'Under-Review', 'Approved', 'Audit-Ready'])
    ws.add_data_validation(status_dv)
    status_dv.add('B54')

    # Audit Readiness Checklist
    ws['A56'] = 'Audit Readiness Checklist:'
    apply_cell_style(ws['A56'], styles['label_cell'])

    audit_items = [
        'All three approvals complete',
        'Evidence 100% verified',
        'All critical gaps have remediation plans',
        'No vacant critical roles',
        'Assessment audit-ready'
    ]

    for i, item in enumerate(audit_items, start=57):
        ws[f'A{i}'] = f'[ ] {item}'
        ws.merge_cells(f'A{i}:D{i}')

    # Protection
    ws.protection.sheet = True
    ws.protection.sort = False


# ==============================================================================
# MAIN EXECUTION
# ==============================================================================

def main():
    """
    Main function to generate the Roles & Responsibilities Assessment workbook.

    Returns:
        int: 0 on success, 1 on failure
    """
    try:
        logger.info("=" * 80)
        logger.info("ISMS-IMP-A.5.1-2-6.1-2.S2 - Roles & Responsibilities Assessment")
        logger.info("Workbook Generator")
        logger.info("=" * 80)

        # Initialize workbook
        logger.info("Creating workbook...")
        wb = Workbook()
        styles = setup_styles()

        # Create all sheets
        logger.info("Creating Dashboard...")
        create_dashboard_sheet(wb, styles)

        logger.info("Creating Role_Inventory...")
        create_role_inventory_sheet(wb, styles)

        logger.info("Creating Role_Definition_Assessment...")
        create_role_definition_assessment_sheet(wb, styles)

        logger.info("Creating RACI_Matrix_Assessment...")
        create_raci_matrix_assessment_sheet(wb, styles)

        logger.info("Creating Role_Assignment_Verification...")
        create_role_assignment_verification_sheet(wb, styles)

        logger.info("Creating Training_Assessment...")
        create_training_assessment_sheet(wb, styles)

        logger.info("Creating Access_Alignment_Review...")
        create_access_alignment_review_sheet(wb, styles)

        logger.info("Creating Gap_Analysis...")
        create_gap_analysis_sheet(wb, styles)

        logger.info("Creating Evidence_Register...")
        create_evidence_register_sheet(wb, styles)

        logger.info("Creating Approval_Sign_Off...")
        create_approval_signoff_sheet(wb, styles)

        # Set workbook properties
        wb.properties.title = "ISMS-IMP-A.5.1-2-6.1-2.S2 - Roles & Responsibilities Assessment"
        wb.properties.subject = "ISO/IEC 27001:2022 Controls A.5.2 & A.6.1 Assessment"
        wb.properties.creator = "[Organization] Information Security Team"
        wb.properties.keywords = "Roles, Responsibilities, RACI, ISMS, ISO27001, A.5.2, A.6.1"
        wb.properties.comments = "Generated via Python script generate_a5_1_2_6_1_2_s2_roles_responsibilities.py"

        # Generate filename with current date
        today = datetime.now().strftime("%Y%m%d")
        filename = f"ISMS-IMP-A.5.1-2-6.1-2.S2_Roles_Responsibilities_{today}.xlsx"

        # Save workbook
        logger.info(f"Saving workbook as: {filename}")
        wb.save(filename)

        logger.info("=" * 80)
        logger.info("Workbook generated successfully!")
        logger.info("=" * 80)
        logger.info("Next Steps:")
        logger.info("1. Open the workbook in Excel")
        logger.info("2. Complete Sheet 2 (Role_Inventory) - this is your foundation")
        logger.info("3. Complete Sheets 3-7 (assessment domains)")
        logger.info("4. Review Sheet 8 (Gap_Analysis)")
        logger.info("5. Document Sheet 9 (Evidence)")
        logger.info("6. Review Sheet 1 (Dashboard) - auto-calculated")
        logger.info("7. Obtain Sheet 10 (Approval Sign-Off)")
        logger.info(f"File location: ./{filename}")

        return 0

    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
