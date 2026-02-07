#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
Excel Workbook Sanity Checker - ISMS A.5.1-2-6.1-2 Secure Employment and Roles
================================================================================

Domain-specific diagnostic utility for A.5.1-2-6.1-2 Secure Employment and Roles
assessment workbooks.

**Purpose:**
Validates workbook structure, formulas, data validations, and cross-references
for all five component workbooks in the Secure Employment and Roles control set.

**Usage:**
    python3 sanity_check_a5_1_2_6_1_2.py [workbook.xlsx]
    python3 sanity_check_a5_1_2_6_1_2.py --all

**Workbooks Validated:**
- S1: Policy Framework Assessment (11 sheets)
- S2: Roles & Responsibilities (10 sheets)
- S3: Screening & Vetting (10 sheets)
- S4: Employment Contract (10 sheets)
- S5: Governance Dashboard (11 sheets)

**Checks Performed:**
- Sheet structure validation
- Data validation dropdowns
- Formula integrity
- Cross-reference linkage
- Conditional formatting presence
- Evidence register structure

**Exit Codes:**
    0 = All checks passed
    1 = Warnings detected (workbook usable)
    2 = Critical errors detected (regenerate recommended)

Control Reference: ISO/IEC 27001:2022 Annex A Controls A.5.1, A.5.2, A.6.1, A.6.2
Assessment Type: Secure Employment and Roles Comprehensive Assessment
Version: 1.0
================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
import os
import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# WORKBOOK DEFINITIONS
# =============================================================================
WORKBOOK_DEFINITIONS = {
    'S1': {
        'name': 'Policy Framework Assessment',
        'pattern': '*S1_Policy_Framework*.xlsx',
        'expected_sheets': [
            'Dashboard',
            'Policy_Inventory',
            'Lifecycle_Compliance',
            'Governance_Assessment',
            'Classification_Review',
            'Communication_Tracking',
            'Repository_Assessment',
            'Gap_Analysis',
            'Evidence_Register',
            'Action_Items',
            'Approval_Sign_Off'
        ],
        'key_validations': {
            'Policy_Status': ['Draft', 'Under Review', 'Approved', 'Retired', 'Archived'],
            'Classification': ['Public', 'Internal', 'Confidential', 'Restricted'],
            'Compliance_Status': ['Compliant', 'Partially Compliant', 'Non-Compliant', 'Not Assessed']
        }
    },
    'S2': {
        'name': 'Roles & Responsibilities',
        'pattern': '*S2_Roles_Responsibilities*.xlsx',
        'expected_sheets': [
            'Dashboard',
            'Role_Inventory',
            'Role_Definition_Assessment',
            'RACI_Matrix_Assessment',
            'Role_Assignment_Verification',
            'Training_Assessment',
            'Access_Alignment_Review',
            'Gap_Analysis',
            'Evidence_Register',
            'Approval_Sign_Off'
        ],
        'key_validations': {
            'RACI': ['Responsible', 'Accountable', 'Consulted', 'Informed', 'R', 'A', 'C', 'I'],
            'Role_Status': ['Active', 'Inactive', 'Pending', 'Under Review'],
            'Assignment_Status': ['Assigned', 'Vacant', 'Temporary', 'Shared']
        }
    },
    'S3': {
        'name': 'Screening & Vetting',
        'pattern': '*S3_Screening_Vetting*.xlsx',
        'expected_sheets': [
            'Dashboard',
            'Screening_Process_Assessment',
            'Screening_Level_Matrix',
            'Personnel_Screening_Registry',
            'Screening_Compliance_Verif',
            'Continuous_Screening_Assessment',
            'Legal_Compliance_Review',
            'Gap_Analysis',
            'Evidence_Register',
            'Approval_Sign_Off'
        ],
        'key_validations': {
            'Screening_Level': ['Basic', 'Standard', 'Enhanced', 'Deep', 'None Required'],
            'Screening_Status': ['Completed', 'In Progress', 'Pending', 'Expired', 'Waived'],
            'Verification_Result': ['Passed', 'Failed', 'Pending Review', 'Conditional']
        }
    },
    'S4': {
        'name': 'Employment Contract',
        'pattern': '*S4_Employment_Contract*.xlsx',
        'expected_sheets': [
            'Dashboard',
            'Contract_Template_Assessment',
            'Required_Clause_Registry',
            'Personnel_Contract_Compliance',
            'Confidentiality_NDA_Tracking',
            'Post_Employment_Obligations',
            'Contractor_Agreement_Assessment',
            'Gap_Analysis',
            'Evidence_Register',
            'Approval_Sign_Off'
        ],
        'key_validations': {
            'Contract_Status': ['Active', 'Expired', 'Pending', 'Terminated'],
            'Clause_Compliance': ['Included', 'Missing', 'Partial', 'N/A'],
            'NDA_Status': ['Signed', 'Pending', 'Expired', 'Not Required']
        }
    },
    'S5': {
        'name': 'Governance Dashboard',
        'pattern': '*S5_Governance*.xlsx',
        'expected_sheets': [
            'Instructions_Legend',
            'Executive_Summary',
            'Maturity_Assessment',
            'Policy_Summary',
            'Roles_Summary',
            'Screening_Summary',
            'Contract_Summary',
            'Gap_Analysis',
            'Evidence_Register',
            'Action_Items',
            'Approval_Sign_Off'
        ],
        'key_validations': {
            'Maturity_Level': ['Initial', 'Managed', 'Defined', 'Quantitatively Managed', 'Optimizing',
                               '1', '2', '3', '4', '5'],
            'Overall_Status': ['Green', 'Yellow', 'Red', 'On Track', 'At Risk', 'Critical'],
            'Priority': ['Critical', 'High', 'Medium', 'Low']
        }
    }
}


def check_sheet_structure(wb, expected_sheets, workbook_id):
    """Validate that all expected sheets are present."""
    issues = []
    warnings = []

    found_sheets = wb.sheetnames
    missing_sheets = [s for s in expected_sheets if s not in found_sheets]
    extra_sheets = [s for s in found_sheets if s not in expected_sheets]

    if missing_sheets:
        for sheet in missing_sheets:
            issues.append(f"  [X] Missing required sheet: '{sheet}'")

    if extra_sheets:
        for sheet in extra_sheets:
            warnings.append(f"  [!] Unexpected sheet found: '{sheet}'")

    logger.info(f"  Expected: {len(expected_sheets)} sheets")
    logger.info(f"  Found: {len(found_sheets)} sheets")

    if not missing_sheets and not extra_sheets:
        logger.info("  [OK] All expected sheets present, no extras")

    return issues, warnings


def check_data_validations(wb, key_validations, workbook_id):
    """Check for expected data validation dropdowns."""
    issues = []
    warnings = []

    total_validations = 0
    validations_found = {key: False for key in key_validations.keys()}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, 'data_validations') and ws.data_validations:
            dv_count = len(ws.data_validations.dataValidation)
            total_validations += dv_count

            for dv in ws.data_validations.dataValidation:
                if dv.formula1:
                    formula = str(dv.formula1)
                    for key, values in key_validations.items():
                        if any(v in formula for v in values):
                            validations_found[key] = True

    logger.info(f"  Total data validations found: {total_validations}")

    for key, found in validations_found.items():
        if found:
            logger.info(f"  [OK] {key} validation detected")
        else:
            warnings.append(f"  [!] {key} validation not detected")

    if total_validations == 0:
        warnings.append("  [!] No data validations found in workbook")

    return issues, warnings


def check_formula_integrity(wb, workbook_id):
    """Check formulas for syntax errors."""
    issues = []
    warnings = []

    formula_count = 0
    formula_errors = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                try:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
                        formula = cell.value

                        # Check for common errors
                        if formula.count('(') != formula.count(')'):
                            issues.append(f"  [X] {sheet_name}!{cell.coordinate}: Unbalanced parentheses")
                            formula_errors += 1

                        if formula.count('"') % 2 != 0:
                            issues.append(f"  [X] {sheet_name}!{cell.coordinate}: Unbalanced quotes")
                            formula_errors += 1

                        # Check for #REF! style errors in formula text
                        if '#REF!' in formula or '#NAME?' in formula:
                            issues.append(f"  [X] {sheet_name}!{cell.coordinate}: Contains error reference")
                            formula_errors += 1
                except Exception as e:  # Skip merged/protected cells
                    pass

    logger.info(f"  Total formulas found: {formula_count}")

    if formula_errors == 0:
        logger.info("  [OK] No formula syntax errors detected")
    else:
        logger.info(f"  [X] Formula errors found: {formula_errors}")

    return issues, warnings


def check_conditional_formatting(wb, workbook_id):
    """Check for presence of conditional formatting."""
    issues = []
    warnings = []

    total_cf_rules = 0
    sheets_with_cf = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
            cf_count = len(ws.conditional_formatting._cf_rules)
            if cf_count > 0:
                sheets_with_cf += 1
                total_cf_rules += cf_count

    logger.info(f"  Total conditional formatting rules: {total_cf_rules}")
    logger.info(f"  Sheets with conditional formatting: {sheets_with_cf}")

    if total_cf_rules > 0:
        logger.info("  [OK] Conditional formatting present")
    else:
        warnings.append("  [!] No conditional formatting detected")

    return issues, warnings


def check_evidence_register(wb, workbook_id):
    """Check Evidence Register structure and content."""
    issues = []
    warnings = []

    evidence_sheet_names = ['Evidence_Register', 'Evidence Register']
    evidence_sheet = None

    for name in evidence_sheet_names:
        if name in wb.sheetnames:
            evidence_sheet = name
            break

    if evidence_sheet:
        ws = wb[evidence_sheet]

        has_content = False
        for row in ws.iter_rows(min_row=1, max_row=15, max_col=10):
            for cell in row:
                if cell.value:
                    has_content = True
                    break
            if has_content:
                break

        if has_content:
            logger.info("  [OK] Evidence Register has content")
        else:
            warnings.append("  [!] Evidence Register appears empty")

        # Check for data validations
        if hasattr(ws, 'data_validations') and ws.data_validations:
            logger.info(f"  [OK] Evidence Register has {len(ws.data_validations.dataValidation)} validations")
        else:
            warnings.append("  [!] No data validations in Evidence Register")
    else:
        issues.append("  [X] Evidence Register sheet not found")

    return issues, warnings


def check_gap_analysis(wb, workbook_id):
    """Check Gap Analysis structure."""
    issues = []
    warnings = []

    gap_sheet_names = ['Gap_Analysis', 'Gap Analysis']
    gap_sheet = None

    for name in gap_sheet_names:
        if name in wb.sheetnames:
            gap_sheet = name
            break

    if gap_sheet:
        ws = wb[gap_sheet]

        expected_severities = ['Critical', 'High', 'Medium', 'Low']
        expected_statuses = ['Open', 'In Progress', 'Planned', 'Resolved', 'Closed']

        if hasattr(ws, 'data_validations') and ws.data_validations:
            logger.info(f"  [OK] Gap Analysis has {len(ws.data_validations.dataValidation)} validations")

            severity_found = False
            status_found = False

            for dv in ws.data_validations.dataValidation:
                if dv.formula1:
                    formula = str(dv.formula1)
                    if any(s in formula for s in expected_severities):
                        severity_found = True
                    if any(s in formula for s in expected_statuses):
                        status_found = True

            if severity_found:
                logger.info("  [OK] Severity validation detected")
            else:
                warnings.append("  [!] Severity validation not detected")

            if status_found:
                logger.info("  [OK] Status validation detected")
            else:
                warnings.append("  [!] Status validation not detected")
        else:
            warnings.append("  [!] No data validations in Gap Analysis")
    else:
        issues.append("  [X] Gap Analysis sheet not found")

    return issues, warnings


def check_cross_references(wb, workbook_id):
    """Check for cross-sheet references in formulas."""
    issues = []
    warnings = []

    cross_ref_count = 0
    sheets_referenced = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                try:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula = cell.value
                        # Check for sheet references (Sheet!Cell or 'Sheet Name'!Cell)
                        if '!' in formula:
                            cross_ref_count += 1
                            # Extract referenced sheet name
                            if "'" in formula:
                                # Handle quoted sheet names
                                parts = formula.split("'")
                                for i, part in enumerate(parts):
                                    if i % 2 == 1:  # Odd indices are sheet names
                                        sheets_referenced.add(part)
                            else:
                                # Handle unquoted sheet names
                                for part in formula.split('!'):
                                    if '(' in part:
                                        ref = part.split('(')[-1]
                                        if ref and ref[0].isalpha():
                                            sheets_referenced.add(ref)
                except Exception as e:  # Skip merged/protected cells
                    pass

    if cross_ref_count > 0:
        logger.info(f"  [OK] Cross-sheet references found: {cross_ref_count}")
        logger.info(f"  Referenced sheets: {len(sheets_referenced)}")
    else:
        logger.info("  [INFO] No cross-sheet references detected")

    return issues, warnings


def check_single_workbook(filename, workbook_def=None):
    """Validate a single workbook against its definition."""
    all_issues = []
    all_warnings = []

    # Determine workbook type from filename if not provided
    if workbook_def is None:
        for wb_id, wb_def in WORKBOOK_DEFINITIONS.items():
            if wb_id in filename:
                workbook_def = wb_def
                workbook_id = wb_id
                break
        if workbook_def is None:
            # Try to match by pattern
            for wb_id, wb_def in WORKBOOK_DEFINITIONS.items():
                pattern_base = wb_def['pattern'].replace('*', '').replace('.xlsx', '')
                if pattern_base in filename:
                    workbook_def = wb_def
                    workbook_id = wb_id
                    break
    else:
        workbook_id = 'UNKNOWN'
        for wb_id, wb_def in WORKBOOK_DEFINITIONS.items():
            if wb_def == workbook_def:
                workbook_id = wb_id
                break

    if workbook_def is None:
        logger.error(f"Could not determine workbook type for: {filename}")
        return [], ["Could not determine workbook type"]

    logger.info("=" * 80)
    logger.info(f"VALIDATING: {workbook_id} - {workbook_def['name']}")
    logger.info("=" * 80)
    logger.info(f"File: {os.path.basename(filename)}")

    # Load workbook
    try:
        wb = load_workbook(filename, data_only=False)
        logger.info(f"[OK] Workbook loaded: {len(wb.sheetnames)} sheets")
    except Exception as e:
        logger.error(f"[X] CRITICAL: Cannot load workbook: {e}")
        return [f"Cannot load workbook: {e}"], []

    # Check 1: Sheet Structure
    logger.info("")
    logger.info("-" * 40)
    logger.info("CHECK 1: SHEET STRUCTURE")
    logger.info("-" * 40)
    issues, warnings = check_sheet_structure(wb, workbook_def['expected_sheets'], workbook_id)
    all_issues.extend(issues)
    all_warnings.extend(warnings)

    # Check 2: Data Validations
    logger.info("")
    logger.info("-" * 40)
    logger.info("CHECK 2: DATA VALIDATIONS")
    logger.info("-" * 40)
    issues, warnings = check_data_validations(wb, workbook_def['key_validations'], workbook_id)
    all_issues.extend(issues)
    all_warnings.extend(warnings)

    # Check 3: Formula Integrity
    logger.info("")
    logger.info("-" * 40)
    logger.info("CHECK 3: FORMULA INTEGRITY")
    logger.info("-" * 40)
    issues, warnings = check_formula_integrity(wb, workbook_id)
    all_issues.extend(issues)
    all_warnings.extend(warnings)

    # Check 4: Conditional Formatting
    logger.info("")
    logger.info("-" * 40)
    logger.info("CHECK 4: CONDITIONAL FORMATTING")
    logger.info("-" * 40)
    issues, warnings = check_conditional_formatting(wb, workbook_id)
    all_issues.extend(issues)
    all_warnings.extend(warnings)

    # Check 5: Evidence Register
    logger.info("")
    logger.info("-" * 40)
    logger.info("CHECK 5: EVIDENCE REGISTER")
    logger.info("-" * 40)
    issues, warnings = check_evidence_register(wb, workbook_id)
    all_issues.extend(issues)
    all_warnings.extend(warnings)

    # Check 6: Gap Analysis
    logger.info("")
    logger.info("-" * 40)
    logger.info("CHECK 6: GAP ANALYSIS")
    logger.info("-" * 40)
    issues, warnings = check_gap_analysis(wb, workbook_id)
    all_issues.extend(issues)
    all_warnings.extend(warnings)

    # Check 7: Cross-References
    logger.info("")
    logger.info("-" * 40)
    logger.info("CHECK 7: CROSS-REFERENCES")
    logger.info("-" * 40)
    issues, warnings = check_cross_references(wb, workbook_id)
    all_issues.extend(issues)
    all_warnings.extend(warnings)

    wb.close()
    return all_issues, all_warnings


def find_workbooks(base_path):
    """Find all workbooks in the 90_workbooks directory."""
    workbooks = {}
    workbook_dir = os.path.join(base_path, '90_workbooks')

    if not os.path.exists(workbook_dir):
        logger.error(f"Workbook directory not found: {workbook_dir}")
        return workbooks

    for wb_id, wb_def in WORKBOOK_DEFINITIONS.items():
        pattern = os.path.join(workbook_dir, wb_def['pattern'])
        matches = glob.glob(pattern)
        if matches:
            # Use the most recent file if multiple exist
            workbooks[wb_id] = max(matches, key=os.path.getmtime)

    return workbooks


def main():
    """Main entry point for sanity check."""
    logger.info("=" * 80)
    logger.info("A.5.1-2-6.1-2 SECURE EMPLOYMENT AND ROLES - SANITY CHECK")
    logger.info("=" * 80)

    if len(sys.argv) < 2:
        logger.info("")
        logger.info("Usage:")
        logger.info("  python3 sanity_check_a5_1_2_6_1_2.py <filename.xlsx>")
        logger.info("  python3 sanity_check_a5_1_2_6_1_2.py --all")
        logger.info("")
        logger.info("Examples:")
        logger.info("  python3 sanity_check_a5_1_2_6_1_2.py ../90_workbooks/ISMS-IMP-A.5.1-2-6.1-2.S1_Policy_Framework_20260131.xlsx")
        logger.info("  python3 sanity_check_a5_1_2_6_1_2.py --all")
        logger.info("")
        logger.info("Workbooks validated:")
        for wb_id, wb_def in WORKBOOK_DEFINITIONS.items():
            logger.info(f"  {wb_id}: {wb_def['name']} ({len(wb_def['expected_sheets'])} sheets)")
        logger.info("")
        logger.info("Exit codes:")
        logger.info("  0 = All checks passed")
        logger.info("  1 = Warnings detected (workbook usable)")
        logger.info("  2 = Critical errors detected (regenerate recommended)")
        logger.info("=" * 80)
        sys.exit(1)

    total_issues = []
    total_warnings = []

    if sys.argv[1] == '--all':
        # Find and validate all workbooks
        script_dir = os.path.dirname(os.path.abspath(__file__))
        base_path = os.path.dirname(script_dir)

        workbooks = find_workbooks(base_path)

        if not workbooks:
            logger.error("No workbooks found to validate")
            sys.exit(2)

        logger.info(f"Found {len(workbooks)} workbooks to validate")
        logger.info("")

        for wb_id, filepath in sorted(workbooks.items()):
            issues, warnings = check_single_workbook(filepath, WORKBOOK_DEFINITIONS[wb_id])
            total_issues.extend(issues)
            total_warnings.extend(warnings)
            logger.info("")
    else:
        # Validate single workbook
        filename = sys.argv[1]
        if not os.path.exists(filename):
            logger.error(f"File not found: {filename}")
            sys.exit(2)

        issues, warnings = check_single_workbook(filename)
        total_issues.extend(issues)
        total_warnings.extend(warnings)

    # Summary Report
    logger.info("=" * 80)
    logger.info("VALIDATION SUMMARY")
    logger.info("=" * 80)

    if total_issues:
        logger.info("")
        logger.info(f"[X] CRITICAL ISSUES: {len(total_issues)}")
        for issue in total_issues:
            logger.info(issue)

    if total_warnings:
        logger.info("")
        logger.info(f"[!] WARNINGS: {len(total_warnings)}")
        for warning in total_warnings:
            logger.info(warning)

    if not total_issues and not total_warnings:
        logger.info("")
        logger.info("[OK] ALL CHECKS PASSED")
        logger.info("")
        logger.info("The A.5.1-2-6.1-2 Secure Employment and Roles workbooks appear healthy.")
        logger.info("Ready for assessment data collection.")
    else:
        logger.info("")
        logger.info("=" * 80)
        logger.info("RECOMMENDATIONS:")
        logger.info("=" * 80)

        if any("Missing required sheet" in i for i in total_issues):
            logger.info("")
            logger.info("- Regenerate workbooks with the appropriate generator script")
            logger.info("  to ensure all required sheets are present")

        if total_warnings:
            logger.info("")
            logger.info("- Review warnings - most are informational")
            logger.info("- Data validations improve data quality but some may be optional")
            logger.info("- Ensure assessment data is current and accurate")

    logger.info("")
    logger.info("=" * 80)

    # Return exit code
    if total_issues:
        sys.exit(2)
    elif total_warnings:
        sys.exit(1)
    else:
        sys.exit(0)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED (syntax validated, structure verified)
# QA_TOOL: Claude Code Deep Scan
# QA_NOTE: STANDARDIZATION - License header, logging, main() pattern applied
# =============================================================================
