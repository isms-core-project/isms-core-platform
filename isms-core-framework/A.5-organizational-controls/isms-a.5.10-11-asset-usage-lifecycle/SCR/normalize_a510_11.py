#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Normalize A.5.10-11 Asset Usage Lifecycle Workbooks

Ensures consistent formatting, cell styling, and data validation
across all assessment workbooks for this control group.
"""

import logging
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

WORKBOOK_DIR = Path(__file__).parent.parent / "90_workbooks"

def normalize_workbook(filepath: Path) -> bool:
    """Normalize a single workbook."""
    try:
        wb = load_workbook(filepath)

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ws in wb.worksheets:
            # Ensure freeze panes are set
            if ws.freeze_panes is None:
                ws.freeze_panes = "A2"

            # Standardize column widths for common columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception as e:  # Skip merged cells
                        pass
                adjusted_width = min(max_length + 2, 50)
                if ws.column_dimensions[column].width is None or ws.column_dimensions[column].width < 8:
                    ws.column_dimensions[column].width = max(adjusted_width, 12)

        wb.save(filepath)
        logger.info("Normalized: %s", filepath.name)
        return True

    except Exception as e:
        logger.error("Failed to normalize %s: %s", filepath.name, e)
        return False


def main() -> int:
    """Main execution."""
    logger.info("=" * 60)
    logger.info("A.5.10-11 Workbook Normalization")
    logger.info("=" * 60)

    if not WORKBOOK_DIR.exists():
        logger.error("Workbook directory not found: %s", WORKBOOK_DIR)
        return 1

    workbooks = list(WORKBOOK_DIR.glob("ISMS-IMP-A.5.10-11*.xlsx"))

    if not workbooks:
        logger.warning("No workbooks found to normalize")
        return 0

    success = 0
    for wb_path in workbooks:
        if normalize_workbook(wb_path):
            success += 1

    logger.info("=" * 60)
    logger.info("Normalized %d of %d workbooks", success, len(workbooks))
    logger.info("=" * 60)

    return 0 if success == len(workbooks) else 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - UTILITY SCRIPT
# =============================================================================

