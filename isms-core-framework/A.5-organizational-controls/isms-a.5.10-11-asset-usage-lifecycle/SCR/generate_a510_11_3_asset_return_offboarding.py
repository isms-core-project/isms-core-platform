#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.10-11.S3 - Asset Return & Offboarding Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.11: Return of Assets

Assessment Domain 3 of 4: Asset Return & Offboarding Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
asset return processes during employment termination, contract completion, or
role changes.

**Control Requirement:**
Personnel and other interested parties as appropriate should return all the
organisation's assets in their possession upon change or termination of their
employment, contract or agreement.

**Generated Workbook Structure:**
1. Instructions - Assessment guidance and methodology
2. Return_Process - Process completeness assessment
3. Asset_Checklist - Standard return checklist by asset type
4. Offboarding_Tracking - Individual offboarding tracking
5. Access_Revocation - Access removal verification
6. Evidence_Register - Audit evidence tracking
7. Approval_SignOff - Stakeholder review and approval

================================================================================
"""

# =============================================================================
# IMPORTS - Standard Library
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# IMPORTS - Third Party
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.10-11.S3"
WORKBOOK_NAME = "Asset Return and Offboarding Assessment"
CONTROL_ID = "A.5.11"
CONTROL_NAME = "Return of Assets"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "returned": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "pending": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "overdue": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Return_Process",
        "Asset_Checklist",
        "Offboarding_Tracking",
        "Access_Revocation",
        "Evidence_Register",
        "Approval_SignOff",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet with assessment guidance."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Asset Return and Offboarding Process"),
        ("Control Reference", CONTROL_ID),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Annual"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "CONTROL REQUIREMENT"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    ws[f"A{row}"] = (
        "ISO 27001:2022 A.5.11 requires that personnel and other interested parties "
        "as appropriate should return all the organisation's assets in their possession "
        "upon change or termination of their employment, contract or agreement."
    )
    ws.merge_cells(f"A{row}:G{row}")

    row += 2
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Assess return process completeness in Return_Process",
        "2. Review standard asset checklist in Asset_Checklist",
        "3. Track individual offboardings in Offboarding_Tracking",
        "4. Verify access revocation in Access_Revocation",
        "5. Link supporting evidence in Evidence_Register",
        "6. Obtain approvals in Approval_SignOff sheet",
    ]

    row += 1
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    row += 1
    ws[f"A{row}"] = "ASSET CATEGORIES FOR RETURN"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    asset_categories = [
        ("Physical Assets", "Laptops, mobile devices, tokens, keys, badges"),
        ("Digital Assets", "Accounts, certificates, encryption keys"),
        ("Information Assets", "Documents, data copies, project files"),
        ("Authentication", "Passwords, PINs, smart cards, biometric records"),
        ("Access Rights", "System access, physical access, VPN, cloud services"),
        ("Third-Party Access", "Partner portals, client systems, vendor accounts"),
    ]

    row += 1
    for category, examples in asset_categories:
        ws[f"A{row}"] = category
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = examples
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    ws.freeze_panes = "A4"


# =============================================================================
# RETURN PROCESS SHEET
# =============================================================================
def create_return_process_sheet(ws, styles):
    """Create the Return_Process sheet - process assessment."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "ASSET RETURN PROCESS ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Requirement_ID", 14),
        ("Process_Requirement", 45),
        ("Category", 22),
        ("Implemented", 14),
        ("Documented", 14),
        ("Responsible_Role", 22),
        ("SLA_Days", 12),
        ("Automated", 14),
        ("Gap_Status", 16),
        ("Remediation_Notes", 35),
        ("Evidence_Ref", 18),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Process requirements for A.5.11
    requirements = [
        ("PR-001", "Formal offboarding procedure documented", "Process"),
        ("PR-002", "Asset inventory linked to employee records", "Asset Management"),
        ("PR-003", "Return checklist defined for each asset type", "Checklist"),
        ("PR-004", "IT notified upon termination/transfer notice", "Communication"),
        ("PR-005", "HR/Manager triggers offboarding workflow", "Workflow"),
        ("PR-006", "Physical asset return point designated", "Physical"),
        ("PR-007", "Asset condition assessment performed", "Verification"),
        ("PR-008", "Data wiping procedure for returned devices", "Data Security"),
        ("PR-009", "Access revocation within 24 hours of last day", "Access Control"),
        ("PR-010", "Knowledge transfer process defined", "Knowledge"),
        ("PR-011", "Third-party asset return coordinated", "Third Party"),
        ("PR-012", "BYOD data removal verified", "BYOD"),
        ("PR-013", "Remote employee return shipping process", "Remote"),
        ("PR-014", "Incomplete return escalation process", "Escalation"),
        ("PR-015", "Return completion sign-off obtained", "Sign-off"),
        ("PR-016", "Asset inventory updated post-return", "Inventory"),
        ("PR-017", "Exceptions documented and approved", "Exceptions"),
        ("PR-018", "Audit trail maintained for all returns", "Audit"),
    ]

    dv_implemented = DataValidation(
        type="list",
        formula1='"Yes,Partial,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_implemented)

    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_yesno)

    dv_gap = DataValidation(
        type="list",
        formula1='"Compliant,Gap Identified,Remediation In Progress"',
        allow_blank=False
    )
    ws.add_data_validation(dv_gap)

    for row_idx, (req_id, req, category) in enumerate(requirements, start=4):
        ws.cell(row=row_idx, column=1, value=req_id)
        ws.cell(row=row_idx, column=2, value=req)
        ws.cell(row=row_idx, column=3, value=category)

        for c in range(4, 12):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_implemented.add(ws.cell(row=row_idx, column=4))
        dv_yesno.add(ws.cell(row=row_idx, column=5))
        dv_yesno.add(ws.cell(row=row_idx, column=8))
        dv_gap.add(ws.cell(row=row_idx, column=9))

    # Additional blank rows
    for r in range(len(requirements) + 4, len(requirements) + 14):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_implemented.add(ws.cell(row=r, column=4))
        dv_yesno.add(ws.cell(row=r, column=5))
        dv_yesno.add(ws.cell(row=r, column=8))
        dv_gap.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "C4"


# =============================================================================
# ASSET CHECKLIST SHEET
# =============================================================================
def create_asset_checklist_sheet(ws, styles):
    """Create the Asset_Checklist sheet - standard return checklist."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "STANDARD ASSET RETURN CHECKLIST"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Asset_Type", 25),
        ("Description", 40),
        ("Category", 18),
        ("Return_Required", 16),
        ("Verification_Method", 25),
        ("Responsible_Team", 20),
        ("SLA_Days", 12),
        ("Data_Wipe_Required", 16),
        ("Disposal_If_Not_Returned", 25),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate standard asset types
    asset_types = [
        ("Laptop/Desktop", "Company-issued computer equipment", "Hardware", "Yes"),
        ("Mobile Phone", "Company-issued mobile device", "Hardware", "Yes"),
        ("Tablet", "Company-issued tablet device", "Hardware", "Yes"),
        ("Monitor", "External display equipment", "Hardware", "Yes"),
        ("Docking Station", "Laptop docking equipment", "Hardware", "Yes"),
        ("Keyboard/Mouse", "Input peripherals", "Hardware", "Optional"),
        ("Headset", "Audio communication equipment", "Hardware", "Optional"),
        ("USB Tokens", "Authentication hardware tokens", "Authentication", "Yes"),
        ("Smart Cards", "Physical access/authentication cards", "Authentication", "Yes"),
        ("Physical Keys", "Office/facility keys", "Physical Access", "Yes"),
        ("Access Badges", "Building access cards", "Physical Access", "Yes"),
        ("Corporate Credit Card", "Company payment cards", "Financial", "Yes"),
        ("Vehicle Keys", "Company vehicle access", "Physical", "Yes"),
        ("Software Licenses", "Transferable license media", "Software", "Yes"),
        ("Documentation", "Physical documents, manuals", "Information", "Yes"),
        ("Project Files", "Work product and project data", "Information", "Verify"),
        ("Customer Data", "Client-related information", "Information", "Verify"),
        ("VPN Certificates", "Digital certificates for remote access", "Digital", "Revoke"),
        ("Email Access", "Corporate email account", "Digital", "Disable"),
        ("Cloud Accounts", "SaaS application access", "Digital", "Disable"),
        ("Network Access", "LAN/WiFi authentication", "Digital", "Disable"),
        ("System Accounts", "Application and system logins", "Digital", "Disable"),
    ]

    dv_required = DataValidation(
        type="list",
        formula1='"Yes,No,Optional,Verify"',
        allow_blank=False
    )
    ws.add_data_validation(dv_required)

    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_yesno)

    for row_idx, (asset, desc, category, required) in enumerate(asset_types, start=4):
        ws.cell(row=row_idx, column=1, value=asset)
        ws.cell(row=row_idx, column=2, value=desc)
        ws.cell(row=row_idx, column=3, value=category)
        ws.cell(row=row_idx, column=4, value=required)

        for c in range(5, 11):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_required.add(ws.cell(row=row_idx, column=4))
        dv_yesno.add(ws.cell(row=row_idx, column=8))

    # Additional blank rows
    for r in range(len(asset_types) + 4, len(asset_types) + 14):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_required.add(ws.cell(row=r, column=4))
        dv_yesno.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "C4"


# =============================================================================
# OFFBOARDING TRACKING SHEET
# =============================================================================
def create_offboarding_tracking_sheet(ws, styles):
    """Create the Offboarding_Tracking sheet - individual tracking."""
    ws.merge_cells("A1:N1")
    ws["A1"] = "OFFBOARDING TRACKING REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Offboard_ID", 14),
        ("Employee_Name", 25),
        ("Department", 20),
        ("Employee_Type", 16),
        ("Last_Working_Day", 16),
        ("Offboard_Reason", 18),
        ("Assets_Assigned", 14),
        ("Assets_Returned", 14),
        ("Return_Status", 16),
        ("Access_Revoked", 14),
        ("Data_Wiped", 14),
        ("Sign_Off_Date", 14),
        ("HR_Confirmed", 14),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Employee,Contractor,Consultant,Intern,Temp"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_reason = DataValidation(
        type="list",
        formula1='"Resignation,Termination,Contract End,Transfer,Retirement,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_reason)

    dv_status = DataValidation(
        type="list",
        formula1='"Complete,In Progress,Pending,Overdue"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No,Partial,N/A"',
        allow_blank=False
    )
    ws.add_data_validation(dv_yesno)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"OFF-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 15):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=4))
        dv_reason.add(ws.cell(row=r, column=6))
        dv_status.add(ws.cell(row=r, column=9))
        dv_yesno.add(ws.cell(row=r, column=10))
        dv_yesno.add(ws.cell(row=r, column=11))
        dv_yesno.add(ws.cell(row=r, column=13))

    ws.freeze_panes = "C4"


# =============================================================================
# ACCESS REVOCATION SHEET
# =============================================================================
def create_access_revocation_sheet(ws, styles):
    """Create the Access_Revocation sheet - access removal verification."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "ACCESS REVOCATION VERIFICATION"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Revocation_ID", 14),
        ("Employee_Ref", 18),
        ("Access_Type", 22),
        ("System_Application", 28),
        ("Last_Working_Day", 16),
        ("Revocation_Date", 16),
        ("SLA_Met", 12),
        ("Revoked_By", 22),
        ("Verification_Method", 22),
        ("Verified_By", 22),
        ("Verification_Date", 16),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_access_type = DataValidation(
        type="list",
        formula1='"Network,VPN,Email,Application,Database,Cloud,Physical,Admin"',
        allow_blank=False
    )
    ws.add_data_validation(dv_access_type)

    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_yesno)

    dv_method = DataValidation(
        type="list",
        formula1='"AD Query,System Log,Login Attempt,Screenshot,Attestation"',
        allow_blank=False
    )
    ws.add_data_validation(dv_method)

    for r in range(4, 204):
        ws.cell(row=r, column=1, value=f"REV-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_access_type.add(ws.cell(row=r, column=3))
        dv_yesno.add(ws.cell(row=r, column=7))
        dv_method.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "C4"


# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================
def create_evidence_register_sheet(ws, styles):
    """Create the Evidence_Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 45),
        ("Related_Offboarding", 20),
        ("Collection_Date", 16),
        ("Location", 40),
        ("Collected_By", 25),
        ("Valid_Until", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Sign-off Form,AD Export,System Log,Screenshot,Email Confirmation,Attestation,Checklist"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "C4"


# =============================================================================
# APPROVAL SIGN-OFF SHEET
# =============================================================================
def create_approval_signoff_sheet(ws, styles):
    """Create the Approval_SignOff sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Process Requirements Assessed", "=COUNTA(Return_Process!A4:A31)-COUNTBLANK(Return_Process!B4:B31)"),
        ("Process Requirements Implemented", '=COUNTIF(Return_Process!D4:D31,"Yes")'),
        ("Offboardings Tracked", "=COUNTA(Offboarding_Tracking!A4:A103)-COUNTBLANK(Offboarding_Tracking!B4:B103)"),
        ("Offboardings Complete", '=COUNTIF(Offboarding_Tracking!I4:I103,"Complete")'),
        ("Access Revocations Verified", "=COUNTA(Access_Revocation!A4:A203)-COUNTBLANK(Access_Revocation!B4:B203)"),
        ("Assessment Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "" or (isinstance(value, str) and not value.startswith("=")):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Role/Title", "Department", "Email", "Date"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (IT Manager / HR Manager)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Date", "Signature"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Date", "Approval Decision", "Signature"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add(ws[f"B{row-2}"])

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A3"


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/7] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/7] Creating Return_Process sheet...")
        create_return_process_sheet(wb["Return_Process"], styles)

        logger.info("[3/7] Creating Asset_Checklist sheet...")
        create_asset_checklist_sheet(wb["Asset_Checklist"], styles)

        logger.info("[4/7] Creating Offboarding_Tracking sheet...")
        create_offboarding_tracking_sheet(wb["Offboarding_Tracking"], styles)

        logger.info("[5/7] Creating Access_Revocation sheet...")
        create_access_revocation_sheet(wb["Access_Revocation"], styles)

        logger.info("[6/7] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[7/7] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# END OF SCRIPT
# =============================================================================
# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.10-11.S3 specification
# =============================================================================
