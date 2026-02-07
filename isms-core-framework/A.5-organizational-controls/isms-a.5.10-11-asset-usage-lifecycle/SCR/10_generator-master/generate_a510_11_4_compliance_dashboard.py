#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.10-11.S4 - Asset Usage Lifecycle Compliance Dashboard Generator
================================================================================

ISO/IEC 27001:2022 Controls A.5.10-11: Asset Usage Lifecycle

Assessment Domain 4 of 4: Compliance Dashboard

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel dashboard workbook for monitoring
overall compliance with both A.5.10 (Acceptable Use) and A.5.11 (Return of
Assets) controls.

**Generated Workbook Structure:**
1. Instructions - Dashboard guidance
2. Executive_Summary - High-level compliance overview
3. A510_Compliance - Acceptable Use compliance metrics
4. A511_Compliance - Return of Assets compliance metrics
5. Gap_Register - Consolidated gap tracking
6. Remediation_Tracker - Remediation action tracking
7. Trend_Analysis - Compliance trending over time
8. Approval_SignOff - Stakeholder review and approval

================================================================================
"""

# =============================================================================
# IMPORTS - Standard Library
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# IMPORTS - Third Party
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.10-11.S4"
WORKBOOK_NAME = "Asset Usage Lifecycle Compliance Dashboard"
CONTROL_ID = "A.5.10-11"
CONTROL_NAME = "Asset Usage Lifecycle"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "kpi_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "kpi_value": {
            "font": Font(name="Calibri", size=18, bold=True),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
    }
    return styles


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Executive_Summary",
        "A510_Compliance",
        "A511_Compliance",
        "Gap_Register",
        "Remediation_Tracker",
        "Trend_Analysis",
        "Approval_SignOff",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet with dashboard guidance."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Asset Usage Lifecycle Compliance Dashboard"),
        ("Control Reference", CONTROL_ID),
        ("Version", "1.0"),
        ("Dashboard Date", ""),
        ("Prepared By", ""),
        ("Organisation", ""),
        ("Reporting Period", ""),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "CONTROLS COVERED"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    controls = [
        ("A.5.10", "Acceptable Use of Information and Other Associated Assets"),
        ("A.5.11", "Return of Assets"),
    ]

    row += 1
    for ctrl_id, ctrl_name in controls:
        ws[f"A{row}"] = ctrl_id
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = ctrl_name
        row += 1

    row += 1
    ws[f"A{row}"] = "HOW TO USE THIS DASHBOARD"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Review Executive_Summary for high-level compliance status",
        "2. Drill into A510_Compliance for acceptable use metrics",
        "3. Drill into A511_Compliance for asset return metrics",
        "4. Review Gap_Register for consolidated gaps",
        "5. Track remediation progress in Remediation_Tracker",
        "6. Monitor trends over time in Trend_Analysis",
        "7. Obtain sign-off in Approval_SignOff",
    ]

    row += 1
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    row += 1
    ws[f"A{row}"] = "COMPLIANCE SCORING"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    scoring = [
        ("90-100%", "Compliant - Green", "Controls effectively implemented"),
        ("70-89%", "Partial - Yellow", "Controls partially implemented, improvements needed"),
        ("0-69%", "Non-Compliant - Red", "Significant gaps requiring immediate attention"),
    ]

    row += 1
    for score, rating, desc in scoring:
        ws[f"A{row}"] = score
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = rating
        ws[f"C{row}"] = desc
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 45
    ws.freeze_panes = "A4"


# =============================================================================
# EXECUTIVE SUMMARY SHEET
# =============================================================================
def create_executive_summary_sheet(ws, styles):
    """Create the Executive_Summary sheet - high-level overview."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EXECUTIVE SUMMARY - Asset Usage Lifecycle Compliance"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Overall Compliance Score
    ws.merge_cells("A3:B3")
    ws["A3"] = "OVERALL COMPLIANCE"
    ws["A3"].font = styles["kpi_header"]["font"]
    ws["A3"].fill = styles["kpi_header"]["fill"]
    ws["A3"].alignment = styles["kpi_header"]["alignment"]

    ws.merge_cells("A4:B5")
    ws["A4"] = ""  # Formula to be entered
    ws["A4"].font = styles["kpi_value"]["font"]
    ws["A4"].alignment = styles["kpi_value"]["alignment"]
    ws["A4"].fill = styles["input_cell"]["fill"]
    ws["A4"].border = styles["border"]
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 30

    # A.5.10 Score
    ws.merge_cells("D3:E3")
    ws["D3"] = "A.5.10 COMPLIANCE"
    ws["D3"].font = styles["kpi_header"]["font"]
    ws["D3"].fill = styles["kpi_header"]["fill"]
    ws["D3"].alignment = styles["kpi_header"]["alignment"]

    ws.merge_cells("D4:E5")
    ws["D4"] = ""
    ws["D4"].font = styles["kpi_value"]["font"]
    ws["D4"].alignment = styles["kpi_value"]["alignment"]
    ws["D4"].fill = styles["input_cell"]["fill"]
    ws["D4"].border = styles["border"]

    # A.5.11 Score
    ws.merge_cells("G3:H3")
    ws["G3"] = "A.5.11 COMPLIANCE"
    ws["G3"].font = styles["kpi_header"]["font"]
    ws["G3"].fill = styles["kpi_header"]["fill"]
    ws["G3"].alignment = styles["kpi_header"]["alignment"]

    ws.merge_cells("G4:H5")
    ws["G4"] = ""
    ws["G4"].font = styles["kpi_value"]["font"]
    ws["G4"].alignment = styles["kpi_value"]["alignment"]
    ws["G4"].fill = styles["input_cell"]["fill"]
    ws["G4"].border = styles["border"]

    # Key Metrics Section
    row = 8
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "KEY METRICS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    metrics = [
        ("AUP Policy Completeness", "", "A.5.10"),
        ("Asset Categories Covered", "", "A.5.10"),
        ("User Acknowledgment Rate", "", "A.5.10"),
        ("Usage Rules Documented", "", "A.5.10"),
        ("Return Process Requirements Met", "", "A.5.11"),
        ("Offboarding Completion Rate", "", "A.5.11"),
        ("Access Revocation SLA Compliance", "", "A.5.11"),
        ("Total Open Gaps", "", "Both"),
        ("Critical Gaps", "", "Both"),
        ("Remediations In Progress", "", "Both"),
    ]

    row += 1
    ws[f"A{row}"] = "Metric"
    ws[f"B{row}"] = "Value"
    ws[f"C{row}"] = "Target"
    ws[f"D{row}"] = "Status"
    ws[f"E{row}"] = "Control"
    for c in range(1, 6):
        ws.cell(row=row, column=c).font = styles["column_header"]["font"]
        ws.cell(row=row, column=c).fill = styles["column_header"]["fill"]
        ws.cell(row=row, column=c).alignment = styles["column_header"]["alignment"]
        ws.cell(row=row, column=c).border = styles["border"]

    dv_status = DataValidation(
        type="list",
        formula1='"On Track,At Risk,Behind"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    row += 1
    for metric, value, control in metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=3).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=5, value=control)
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = styles["border"]
        dv_status.add(ws.cell(row=row, column=4))
        row += 1

    # Key Findings Section
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "KEY FINDINGS AND RECOMMENDATIONS"
    ws[f"A{row}"].font = styles["subheader"]["font"]
    ws[f"A{row}"].fill = styles["subheader"]["fill"]
    ws[f"A{row}"].alignment = styles["subheader"]["alignment"]

    row += 1
    ws[f"A{row}"] = "Finding/Recommendation"
    ws[f"D{row}"] = "Priority"
    ws[f"E{row}"] = "Owner"
    ws[f"F{row}"] = "Due Date"
    for label_col in ["A", "D", "E", "F"]:
        ws[f"{label_col}{row}"].font = styles["column_header"]["font"]
        ws[f"{label_col}{row}"].fill = styles["column_header"]["fill"]
        ws[f"{label_col}{row}"].border = styles["border"]

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_priority)

    for r in range(row + 1, row + 11):
        ws.merge_cells(f"A{r}:C{r}")
        ws[f"A{r}"].fill = styles["input_cell"]["fill"]
        ws[f"A{r}"].border = styles["border"]
        for col in ["D", "E", "F"]:
            ws[f"{col}{r}"].fill = styles["input_cell"]["fill"]
            ws[f"{col}{r}"].border = styles["border"]
        dv_priority.add(ws[f"D{r}"])

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14


# =============================================================================
# A510 COMPLIANCE SHEET
# =============================================================================
def create_a510_compliance_sheet(ws, styles):
    """Create the A510_Compliance sheet - Acceptable Use metrics."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "A.5.10 COMPLIANCE - Acceptable Use of Information and Assets"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Compliance_Area", 35),
        ("Requirement", 45),
        ("Status", 16),
        ("Score", 12),
        ("Max_Score", 12),
        ("Evidence_Ref", 18),
        ("Last_Assessed", 14),
        ("Assessor", 20),
        ("Gap_Notes", 35),
        ("Remediation_Ref", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Compliance areas for A.5.10
    compliance_areas = [
        ("Policy Framework", "AUP exists and is approved"),
        ("Policy Framework", "AUP covers all required content areas"),
        ("Policy Framework", "AUP is version controlled"),
        ("Asset Coverage", "All asset categories addressed"),
        ("Asset Coverage", "Usage rules defined per category"),
        ("Asset Coverage", "Handling requirements documented"),
        ("User Awareness", "AUP communicated to all personnel"),
        ("User Awareness", "Acknowledgment process implemented"),
        ("User Awareness", "Acknowledgment tracked and monitored"),
        ("User Awareness", "Refresher training conducted annually"),
        ("Enforcement", "Violations can be detected"),
        ("Enforcement", "Consequences defined and communicated"),
        ("Enforcement", "Exception process documented"),
        ("Monitoring", "Usage monitoring implemented"),
        ("Monitoring", "Monitoring disclosed to users"),
        ("Third Parties", "Contractors acknowledge AUP"),
        ("Third Parties", "Vendor access rules defined"),
        ("Review", "AUP reviewed annually"),
        ("Review", "Updates communicated to users"),
    ]

    dv_status = DataValidation(
        type="list",
        formula1='"Compliant,Partial,Non-Compliant,Not Assessed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for row_idx, (area, requirement) in enumerate(compliance_areas, start=4):
        ws.cell(row=row_idx, column=1, value=area)
        ws.cell(row=row_idx, column=2, value=requirement)

        for c in range(3, 11):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        ws.cell(row=row_idx, column=5, value=1)  # Default max score
        dv_status.add(ws.cell(row=row_idx, column=3))

    # Summary row
    summary_row = len(compliance_areas) + 5
    ws.cell(row=summary_row, column=1, value="TOTAL SCORE").font = Font(bold=True)
    ws.cell(row=summary_row, column=4).value = f"=SUM(D4:D{summary_row-1})"
    ws.cell(row=summary_row, column=5).value = f"=SUM(E4:E{summary_row-1})"
    ws.merge_cells(f"F{summary_row}:G{summary_row}")
    ws.cell(row=summary_row, column=6).value = f'=IF(E{summary_row}>0,D{summary_row}/E{summary_row}*100,"")'
    ws.cell(row=summary_row, column=6).number_format = '0.0"%"'

    for c in range(1, 8):
        ws.cell(row=summary_row, column=c).font = Font(bold=True)
        ws.cell(row=summary_row, column=c).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws.cell(row=summary_row, column=c).border = styles["border"]

    ws.freeze_panes = "C4"


# =============================================================================
# A511 COMPLIANCE SHEET
# =============================================================================
def create_a511_compliance_sheet(ws, styles):
    """Create the A511_Compliance sheet - Return of Assets metrics."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "A.5.11 COMPLIANCE - Return of Assets"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Compliance_Area", 35),
        ("Requirement", 45),
        ("Status", 16),
        ("Score", 12),
        ("Max_Score", 12),
        ("Evidence_Ref", 18),
        ("Last_Assessed", 14),
        ("Assessor", 20),
        ("Gap_Notes", 35),
        ("Remediation_Ref", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Compliance areas for A.5.11
    compliance_areas = [
        ("Process", "Offboarding procedure documented"),
        ("Process", "Asset inventory linked to personnel"),
        ("Process", "Standard return checklist exists"),
        ("Process", "HR/IT integration for notifications"),
        ("Asset Return", "Physical asset return tracked"),
        ("Asset Return", "Digital asset return verified"),
        ("Asset Return", "BYOD data removal confirmed"),
        ("Asset Return", "Return condition assessment performed"),
        ("Access Revocation", "Access removal within SLA"),
        ("Access Revocation", "All system access revoked"),
        ("Access Revocation", "Physical access revoked"),
        ("Access Revocation", "Third-party access removed"),
        ("Data Security", "Data wiping procedure followed"),
        ("Data Security", "Certificates and keys revoked"),
        ("Knowledge Transfer", "Knowledge transfer completed"),
        ("Knowledge Transfer", "Documentation handed over"),
        ("Verification", "Return sign-off obtained"),
        ("Verification", "Audit trail maintained"),
        ("Remote Workers", "Remote asset return shipping"),
        ("Contractors", "Contractor offboarding process"),
    ]

    dv_status = DataValidation(
        type="list",
        formula1='"Compliant,Partial,Non-Compliant,Not Assessed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for row_idx, (area, requirement) in enumerate(compliance_areas, start=4):
        ws.cell(row=row_idx, column=1, value=area)
        ws.cell(row=row_idx, column=2, value=requirement)

        for c in range(3, 11):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        ws.cell(row=row_idx, column=5, value=1)
        dv_status.add(ws.cell(row=row_idx, column=3))

    # Summary row
    summary_row = len(compliance_areas) + 5
    ws.cell(row=summary_row, column=1, value="TOTAL SCORE").font = Font(bold=True)
    ws.cell(row=summary_row, column=4).value = f"=SUM(D4:D{summary_row-1})"
    ws.cell(row=summary_row, column=5).value = f"=SUM(E4:E{summary_row-1})"
    ws.merge_cells(f"F{summary_row}:G{summary_row}")
    ws.cell(row=summary_row, column=6).value = f'=IF(E{summary_row}>0,D{summary_row}/E{summary_row}*100,"")'
    ws.cell(row=summary_row, column=6).number_format = '0.0"%"'

    for c in range(1, 8):
        ws.cell(row=summary_row, column=c).font = Font(bold=True)
        ws.cell(row=summary_row, column=c).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws.cell(row=summary_row, column=c).border = styles["border"]

    ws.freeze_panes = "C4"


# =============================================================================
# GAP REGISTER SHEET
# =============================================================================
def create_gap_register_sheet(ws, styles):
    """Create the Gap_Register sheet - consolidated gap tracking."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "CONSOLIDATED GAP REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Gap_ID", 12),
        ("Control", 12),
        ("Gap_Description", 45),
        ("Gap_Type", 18),
        ("Severity", 14),
        ("Risk_Rating", 14),
        ("Identified_Date", 14),
        ("Owner", 22),
        ("Target_Date", 14),
        ("Status", 16),
        ("Remediation_Ref", 16),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_control = DataValidation(
        type="list",
        formula1='"A.5.10,A.5.11,Both"',
        allow_blank=False
    )
    ws.add_data_validation(dv_control)

    dv_type = DataValidation(
        type="list",
        formula1='"Policy Gap,Process Gap,Technical Gap,Documentation Gap,Training Gap"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_severity = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_severity)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Remediated,Accepted,Closed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"GAP-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_control.add(ws.cell(row=r, column=2))
        dv_type.add(ws.cell(row=r, column=4))
        dv_severity.add(ws.cell(row=r, column=5))
        dv_status.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "D4"


# =============================================================================
# REMEDIATION TRACKER SHEET
# =============================================================================
def create_remediation_tracker_sheet(ws, styles):
    """Create the Remediation_Tracker sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "REMEDIATION ACTION TRACKER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Remediation_ID", 14),
        ("Gap_Ref", 14),
        ("Action_Description", 45),
        ("Control", 12),
        ("Priority", 12),
        ("Owner", 22),
        ("Start_Date", 14),
        ("Target_Date", 14),
        ("Status", 16),
        ("% Complete", 12),
        ("Completion_Date", 14),
        ("Notes", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_priority)

    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,On Hold,Completed,Cancelled"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"REM-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_priority.add(ws.cell(row=r, column=5))
        dv_status.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "D4"


# =============================================================================
# TREND ANALYSIS SHEET
# =============================================================================
def create_trend_analysis_sheet(ws, styles):
    """Create the Trend_Analysis sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "COMPLIANCE TREND ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = [
        ("Assessment_Period", 18),
        ("A510_Score", 14),
        ("A510_Change", 14),
        ("A511_Score", 14),
        ("A511_Change", 14),
        ("Overall_Score", 14),
        ("Open_Gaps", 14),
        ("Notes", 40),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate periods
    periods = ["Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025", "Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"]

    for row_idx, period in enumerate(periods, start=4):
        ws.cell(row=row_idx, column=1, value=period)
        for c in range(2, 9):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Change formulas
        if row_idx > 4:
            ws.cell(row=row_idx, column=3).value = f"=IF(AND(B{row_idx}<>\"\",B{row_idx-1}<>\"\"),B{row_idx}-B{row_idx-1},\"\")"
            ws.cell(row=row_idx, column=5).value = f"=IF(AND(D{row_idx}<>\"\",D{row_idx-1}<>\"\"),D{row_idx}-D{row_idx-1},\"\")"
            ws.cell(row=row_idx, column=3).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            ws.cell(row=row_idx, column=5).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.freeze_panes = "B4"


# =============================================================================
# APPROVAL SIGN-OFF SHEET
# =============================================================================
def create_approval_signoff_sheet(ws, styles):
    """Create the Approval_SignOff sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "DASHBOARD APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Dashboard Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Dashboard Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Reporting Period", ""),
        ("A.5.10 Overall Compliance", ""),
        ("A.5.11 Overall Compliance", ""),
        ("Total Open Gaps", "=COUNTIF(Gap_Register!J4:J53,\"Open\")"),
        ("Remediations In Progress", '=COUNTIF(Remediation_Tracker!I4:I53,"In Progress")'),
        ("Dashboard Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "" or (isinstance(value, str) and not value.startswith("=")):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "PREPARED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Role/Title", "Date"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (Information Security Manager)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Date", "Signature"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    for field in ["Name", "Date", "Approval Decision", "Signature"]:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add(ws[f"B{row-2}"])

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A3"


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/8] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/8] Creating Executive_Summary sheet...")
        create_executive_summary_sheet(wb["Executive_Summary"], styles)

        logger.info("[3/8] Creating A510_Compliance sheet...")
        create_a510_compliance_sheet(wb["A510_Compliance"], styles)

        logger.info("[4/8] Creating A511_Compliance sheet...")
        create_a511_compliance_sheet(wb["A511_Compliance"], styles)

        logger.info("[5/8] Creating Gap_Register sheet...")
        create_gap_register_sheet(wb["Gap_Register"], styles)

        logger.info("[6/8] Creating Remediation_Tracker sheet...")
        create_remediation_tracker_sheet(wb["Remediation_Tracker"], styles)

        logger.info("[7/8] Creating Trend_Analysis sheet...")
        create_trend_analysis_sheet(wb["Trend_Analysis"], styles)

        logger.info("[8/8] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# END OF SCRIPT
# =============================================================================
# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.5.10-11.S4 specification
# =============================================================================
