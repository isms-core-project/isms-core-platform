#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
A.5.5-6 Sanity Check - Validates all A.5.5-6 External Communications workbooks.

Usage: python3 sanity_check_a55_6.py
"""

import logging
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

CONTROL_ID = "A.5.5-6"
WORKBOOK_DIR = Path(__file__).parent.parent / "90_workbooks"

EXPECTED_WORKBOOKS = {
    "A.5.5-6.1": {
        "pattern": "ISMS-IMP-A.5.5-6.S1_Authority_Contacts*.xlsx",
        "required_sheets": ["Instructions", "Authority_Registry", "Contact_Types", "Communication_Log", "Verification_Register", "Evidence_Register", "Approval_SignOff"]
    },
    "A.5.5-6.2": {
        "pattern": "ISMS-IMP-A.5.5-6.S2_Special_Interest*.xlsx",
        "required_sheets": ["Instructions", "Groups_Registry", "Membership_Details", "Engagement_Log", "Intelligence_Received", "Contribution_Log", "Evidence_Register", "Approval_SignOff"]
    },
    "A.5.5-6.3": {
        "pattern": "ISMS-IMP-A.5.5-6.S3_External_Communication*.xlsx",
        "required_sheets": ["Instructions", "Communication_Scenarios", "Notification_Requirements", "Escalation_Matrix", "Approval_Workflow", "Communication_Templates", "Evidence_Register", "Approval_SignOff"]
    },
    "A.5.5-6.4": {
        "pattern": "ISMS-IMP-A.5.5-6.S4_External_Communications*.xlsx",
        "required_sheets": ["Instructions", "Executive_Summary", "Authority_KPIs", "SIG_KPIs", "Compliance_Scorecard", "Gap_Analysis", "Audit_Readiness", "Trend_Analysis", "Evidence_Register", "Approval_SignOff"]
    },
}


def check_workbook(filepath: Path, expected_sheets: list) -> tuple:
    issues = []
    passed = failed = 0
    try:
        wb = load_workbook(filepath, data_only=False)
        for sheet in expected_sheets:
            if sheet in wb.sheetnames:
                passed += 1
            else:
                failed += 1
                issues.append(f"Missing sheet: {sheet}")
        for ws in wb.worksheets:
            if ws["A1"].value is None:
                issues.append(f"Sheet '{ws.title}' has no header")
                failed += 1
            else:
                passed += 1
            if ws.data_validations.dataValidation:
                passed += 1
            elif ws.title != "Instructions":
                issues.append(f"Sheet '{ws.title}' has no data validations")
        wb.close()
    except Exception as e:
        failed += 1
        issues.append(f"Error: {e}")
    return passed, failed, issues


def main() -> int:
    logger.info("=" * 70)
    logger.info(f"{CONTROL_ID} Sanity Check")
    logger.info("=" * 70)

    if not WORKBOOK_DIR.exists():
        logger.warning(f"Workbook directory not found")
        return 0

    total_passed = total_failed = 0
    all_issues = []

    for doc_id, config in EXPECTED_WORKBOOKS.items():
        logger.info(f"\nChecking {doc_id}...")
        workbooks = list(WORKBOOK_DIR.glob(config["pattern"]))
        if not workbooks:
            logger.warning(f"  No workbooks found matching {config['pattern']}")
            total_failed += 1
            all_issues.append(f"{doc_id}: No workbook found")
            continue

        for wb_path in workbooks:
            logger.info(f"  Checking: {wb_path.name}")
            passed, failed, issues = check_workbook(wb_path, config["required_sheets"])
            total_passed += passed
            total_failed += failed
            for issue in issues:
                logger.warning(f"    {issue}")
                all_issues.append(f"{doc_id}: {issue}")
            if not issues:
                logger.info(f"    All checks passed ({passed} checks)")

    logger.info("\n" + "=" * 70)
    logger.info("SANITY CHECK SUMMARY")
    logger.info("=" * 70)
    logger.info(f"Total checks passed: {total_passed}")
    logger.info(f"Total checks failed: {total_failed}")
    if all_issues:
        logger.info(f"\nIssues found ({len(all_issues)}):")
        for issue in all_issues:
            logger.info(f"  - {issue}")
    else:
        logger.info("\n All sanity checks passed!")
    logger.info("=" * 70)
    return 0 if total_failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.5-6 External Communications control
# =============================================================================
