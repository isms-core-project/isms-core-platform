#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.5-6.S3 - External Communication Procedures
================================================================================

ISO/IEC 27001:2022 Controls A.5.5 & A.5.6: External Communications
Assessment Domain 3 of 4: Communication Procedures

This script generates a comprehensive Excel workbook for documenting and
managing procedures for external communications with authorities and
special interest groups per ISO 27001:2022 Controls A.5.5 and A.5.6.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.5-6.S3"
WORKBOOK_NAME = "External Communication Procedures"
CONTROL_ID = "A.5.5-6"
CONTROL_NAME = "Contact with Authorities & Special Interest Groups"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls A.5.5 & A.5.6: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions sheet."""
    ws.title = "Instructions"

    instructions = [
        ["ISMS-IMP-A.5.5-6.S3 - External Communication Procedures"],
        [""],
        ["PURPOSE"],
        ["This workbook documents procedures for external communications with authorities"],
        ["and special interest groups, including escalation paths, approval workflows,"],
        ["and mandatory notification requirements."],
        [""],
        ["SHEETS"],
        ["1. Instructions - This guidance sheet"],
        ["2. Communication_Scenarios - When to contact which external party"],
        ["3. Notification_Requirements - Mandatory notification obligations"],
        ["4. Escalation_Matrix - Internal escalation before external contact"],
        ["5. Approval_Workflow - Who must approve external communications"],
        ["6. Communication_Templates - Standard message templates"],
        ["7. Evidence_Register - Evidence documentation for audits"],
        ["8. Approval_SignOff - Management approval workflow"],
        [""],
        ["KEY SCENARIOS REQUIRING EXTERNAL CONTACT"],
        ["- Data breaches affecting personal data (DPA notification)"],
        ["- Criminal activity or suspected crimes (Law Enforcement)"],
        ["- Regulatory compliance matters (Sector Regulators)"],
        ["- Emerging cyber threats (NCSC, ISACs)"],
        ["- Business disruption (Emergency Services)"],
        ["- Certification audits (Certification Bodies)"],
        [""],
        ["APPROVAL REQUIREMENTS"],
        ["- Routine communications: Department Head approval"],
        ["- Regulatory notifications: Legal Counsel + CISO"],
        ["- Law enforcement contact: CEO/Managing Director"],
        ["- Public statements: CEO + Communications"],
        [""],
        ["DATA VALIDATION"],
        ["- Yellow cells are input fields"],
        ["- Dropdown lists enforce valid values"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SHEETS", "KEY SCENARIOS REQUIRING EXTERNAL CONTACT",
                          "APPROVAL REQUIREMENTS", "DATA VALIDATION"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 90


def create_communication_scenarios_sheet(ws):
    """Create the Communication Scenarios sheet."""
    ws.title = "Communication_Scenarios"

    headers = [
        "Scenario_ID", "Scenario_Name", "Scenario_Category", "Trigger_Event",
        "Primary_Authority", "Secondary_Authority", "SIG_Contact",
        "Response_Time", "Approval_Level", "Internal_Escalation_First",
        "Documentation_Required", "Template_Reference", "Procedure_Steps"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate with standard scenarios
    scenarios = [
        ("SCEN-001", "Personal Data Breach", "Regulatory", "Data breach affecting individuals",
         "Data Protection Authority", "Sector Regulator (if applicable)", "None",
         "72 hours", "DPO + Legal + CISO", "Yes",
         "Breach report, impact assessment, remediation plan", "TPL-DPA-001",
         "1. Contain breach 2. Assess impact 3. Prepare notification 4. Get approvals 5. Notify DPA"),
        ("SCEN-002", "Cyber Crime Incident", "Law Enforcement", "Criminal activity detected",
         "Cantonal Police / Fedpol", "NCSC", "FS-ISAC (if applicable)",
         "Immediate", "CEO + Legal", "Yes",
         "Incident report, evidence preservation log", "TPL-LE-001",
         "1. Preserve evidence 2. Document incident 3. CEO approval 4. Contact police"),
        ("SCEN-003", "Critical Vulnerability", "Threat Intel", "Zero-day or critical CVE",
         "NCSC", "None", "Relevant ISACs",
         "24 hours", "CISO", "Yes",
         "Vulnerability details, impact assessment", "TPL-VULN-001",
         "1. Assess impact 2. Implement mitigations 3. Report to NCSC if requested"),
        ("SCEN-004", "Regulatory Inquiry", "Regulatory", "Formal inquiry received",
         "Requesting Regulator", "None", "None",
         "Per inquiry deadline", "Legal + CEO", "Yes",
         "Inquiry response, supporting evidence", "TPL-REG-001",
         "1. Acknowledge receipt 2. Legal review 3. Prepare response 4. CEO approval"),
        ("SCEN-005", "Physical Security Incident", "Emergency", "Fire, break-in, or physical threat",
         "Emergency Services", "Cantonal Police", "None",
         "Immediate", "Facility Manager / CEO", "No",
         "Incident report, insurance claim", "TPL-PHYS-001",
         "1. Ensure safety 2. Call emergency services 3. Secure premises 4. Document"),
    ]

    row = 2
    for scenario in scenarios:
        for col, value in enumerate(scenario, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Add empty input rows
    for row in range(row, row + 15):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validations
    category_dv = DataValidation(
        type="list",
        formula1='"Regulatory,Law Enforcement,Emergency,Threat Intel,Standards,Other"'
    )
    ws.add_data_validation(category_dv)
    category_dv.add('C2:C50')

    approval_dv = DataValidation(
        type="list",
        formula1='"Department Head,CISO,Legal,DPO + Legal + CISO,CEO + Legal,CEO"'
    )
    ws.add_data_validation(approval_dv)
    approval_dv.add('I2:I50')

    yn_dv = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('J2:J50')

    # Column widths
    widths = [12, 25, 15, 35, 25, 25, 20, 15, 22, 20, 40, 15, 50]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_notification_requirements_sheet(ws):
    """Create the Notification Requirements sheet."""
    ws.title = "Notification_Requirements"

    headers = [
        "Requirement_ID", "Regulation", "Notification_Type", "Authority",
        "Trigger_Condition", "Time_Limit", "Required_Information",
        "Format", "Penalty_for_Non_Compliance", "Internal_Owner",
        "Procedure_Reference", "Last_Review_Date", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate with common requirements
    requirements = [
        ("NOT-001", "GDPR Art. 33", "Data Breach Notification", "Lead Supervisory Authority",
         "Personal data breach likely to result in risk", "72 hours from awareness",
         "Nature of breach, categories affected, likely consequences, measures taken",
         "Online portal or prescribed form", "Up to EUR 10M or 2% global turnover",
         "DPO", "SCEN-001", "", "Applies to all personal data processing"),
        ("NOT-002", "Swiss nDSG Art. 24", "Data Breach Notification", "FDPIC",
         "Personal data breach with high risk to individuals", "As soon as possible",
         "Type of breach, effects, measures taken or planned",
         "Written notification", "Administrative fines up to CHF 250,000",
         "DPO", "SCEN-001", "", "Swiss Federal data protection law"),
        ("NOT-003", "NIS2 Art. 23", "Significant Incident", "CSIRT/Competent Authority",
         "Significant impact on service provision", "24h early warning, 72h notification",
         "Incident details, impact assessment, mitigation measures",
         "As specified by MS", "Up to EUR 10M or 2% global turnover",
         "CISO", "SCEN-003", "", "If in-scope for NIS2"),
        ("NOT-004", "FINMA Circular 2023/1", "Cyber Incident", "FINMA",
         "Critical cyber incident affecting operations", "24 hours",
         "Nature, scope, timeline, measures taken",
         "As per FINMA guidance", "Regulatory sanctions",
         "CISO", "SCEN-002", "", "Swiss financial institutions only"),
    ]

    row = 2
    for req in requirements:
        for col, value in enumerate(req, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Add empty input rows
    for row in range(row, row + 10):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 20, 25, 30, 40, 25, 50, 25, 35, 15, 15, 15, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_escalation_matrix_sheet(ws):
    """Create the Escalation Matrix sheet."""
    ws.title = "Escalation_Matrix"

    headers = [
        "Level", "Scenario_Type", "First_Contact", "Escalation_1",
        "Escalation_2", "Escalation_3", "Time_to_Escalate",
        "External_Contact_Approval", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate escalation matrix
    matrix = [
        ("L1", "Security Incident (Low)", "Security Analyst", "Security Manager",
         "CISO", "CEO", "2 hours", "CISO", ""),
        ("L2", "Security Incident (Medium)", "Security Manager", "CISO",
         "CEO", "Board", "1 hour", "CEO", ""),
        ("L3", "Security Incident (High/Critical)", "CISO", "CEO",
         "Board", "N/A", "30 minutes", "CEO", "Immediate board notification"),
        ("L4", "Data Breach", "DPO", "CISO + Legal",
         "CEO", "Board", "1 hour", "DPO + Legal + CEO", "72h regulatory clock"),
        ("L5", "Criminal Activity", "Security Manager", "CISO",
         "Legal", "CEO", "Immediate", "CEO + Legal", "Preserve evidence"),
        ("L6", "Regulatory Inquiry", "Compliance Officer", "Legal",
         "CEO", "Board", "Per deadline", "Legal + CEO", ""),
        ("L7", "Physical Emergency", "Facility Manager", "HR Director",
         "CEO", "N/A", "Immediate", "Facility Manager", "Safety first"),
    ]

    row = 2
    for item in matrix:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
        row += 1

    # Add empty input rows
    for row in range(row, row + 10):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [8, 30, 20, 20, 15, 12, 18, 25, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_approval_workflow_sheet(ws):
    """Create the Approval Workflow sheet."""
    ws.title = "Approval_Workflow"

    headers = [
        "Workflow_ID", "Communication_Type", "Recipient_Type", "Required_Approvers",
        "Approval_Sequence", "Max_Approval_Time", "Delegate_When_Absent",
        "Documentation_Required", "Post_Communication_Actions"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate workflows
    workflows = [
        ("WF-001", "Routine Information Sharing", "Special Interest Group",
         "Department Head", "Single", "24 hours", "Deputy Head",
         "Email approval", "Log in engagement register"),
        ("WF-002", "Threat Intelligence Sharing", "ISAC / NCSC",
         "CISO", "Single", "4 hours", "Security Manager",
         "Classification review, CISO approval", "Log in contribution register"),
        ("WF-003", "Data Breach Notification", "Data Protection Authority",
         "DPO + Legal Counsel + CISO", "Parallel then CEO",
         "24 hours (within 72h window)", "Deputies for each role",
         "Breach report, impact assessment, approval chain", "Track response, log evidence"),
        ("WF-004", "Law Enforcement Contact", "Police / Fedpol",
         "CEO + Legal Counsel", "Sequential (Legal then CEO)",
         "2 hours", "CFO + External Counsel",
         "Incident report, evidence log, legal review", "Track investigation, update board"),
        ("WF-005", "Regulatory Response", "Any Regulator",
         "Legal Counsel + CEO", "Sequential",
         "Per deadline", "External Counsel + CFO",
         "Response document, supporting evidence", "Track outcome, update compliance"),
        ("WF-006", "Media/Public Statement", "Media / Public",
         "CEO + Communications", "Parallel",
         "As needed", "CFO + External PR",
         "Statement draft, legal review", "Monitor coverage, track response"),
    ]

    row = 2
    for wf in workflows:
        for col, value in enumerate(wf, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Add empty input rows
    for row in range(row, row + 10):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validations
    sequence_dv = DataValidation(
        type="list",
        formula1='"Single,Sequential,Parallel,Parallel then CEO"'
    )
    ws.add_data_validation(sequence_dv)
    sequence_dv.add('E2:E30')

    # Column widths
    widths = [12, 30, 25, 30, 20, 25, 25, 40, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_communication_templates_sheet(ws):
    """Create the Communication Templates sheet."""
    ws.title = "Communication_Templates"

    headers = [
        "Template_ID", "Template_Name", "Purpose", "Recipient_Type",
        "Required_Sections", "Classification", "Review_Date",
        "Owner", "Storage_Location", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate templates
    templates = [
        ("TPL-DPA-001", "Data Breach Notification (DPA)", "Notify data protection authority of breach",
         "Data Protection Authority",
         "Nature of breach, categories affected, approx numbers, consequences, measures taken",
         "Confidential", "", "DPO", "[SharePoint path]", "Use within 72 hours of awareness"),
        ("TPL-LE-001", "Law Enforcement Report", "Report criminal activity to police",
         "Law Enforcement",
         "Incident description, timeline, evidence preserved, point of contact",
         "Confidential", "", "Legal", "[SharePoint path]", "Preserve chain of custody"),
        ("TPL-VULN-001", "Vulnerability Disclosure", "Share vulnerability details with NCSC",
         "Government Cybersecurity",
         "Vulnerability details, affected systems, mitigation status, IOCs",
         "TLP:AMBER", "", "CISO", "[SharePoint path]", "Anonymize as appropriate"),
        ("TPL-REG-001", "Regulatory Response", "Respond to regulatory inquiries",
         "Any Regulator",
         "Reference to inquiry, response summary, supporting evidence list",
         "Confidential", "", "Legal", "[SharePoint path]", "Track all submissions"),
        ("TPL-ISAC-001", "ISAC Threat Report", "Share threat intelligence with ISAC",
         "ISAC",
         "Threat summary, IOCs, TTPs, recommendations",
         "TLP:AMBER", "", "CISO", "[SharePoint path]", "Classification review required"),
    ]

    row = 2
    for tpl in templates:
        for col, value in enumerate(tpl, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Add empty input rows
    for row in range(row, row + 10):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validations
    class_dv = DataValidation(
        type="list",
        formula1='"TLP:RED,TLP:AMBER,TLP:GREEN,TLP:CLEAR,Confidential,Internal,Public"'
    )
    ws.add_data_validation(class_dv)
    class_dv.add('F2:F30')

    # Column widths
    widths = [12, 35, 45, 25, 60, 15, 15, 15, 30, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_evidence_register_sheet(ws):
    """Create the Evidence Register sheet."""
    ws.title = "Evidence_Register"

    headers = [
        "Evidence_ID", "Evidence_Type", "Description", "Related_Procedure",
        "Date_Created", "Created_By", "Storage_Location",
        "Retention_Period", "Review_Date", "Status", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(
        type="list",
        formula1='"Procedure Document,Template,Approval Record,Training Record,Test Record,Other"'
    )
    ws.add_data_validation(type_dv)
    type_dv.add('B2:B100')

    status_dv = DataValidation(type="list", formula1='"Current,Archived,Pending Review"')
    ws.add_data_validation(status_dv)
    status_dv.add('J2:J100')

    # Format input rows
    for row in range(2, 31):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 22, 40, 18, 15, 20, 40, 15, 15, 15, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_approval_signoff_sheet(ws):
    """Create the Approval Sign-Off sheet."""
    ws.title = "Approval_SignOff"

    headers = [
        "Approval_ID", "Review_Period", "Review_Date", "Reviewer_Name",
        "Reviewer_Role", "Procedures_Complete", "Templates_Current",
        "Training_Complete", "Approval_Status", "Signature_Date",
        "Next_Review_Date", "Comments"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    role_dv = DataValidation(
        type="list",
        formula1='"CISO,Legal Counsel,DPO,Compliance Officer,CEO"'
    )
    ws.add_data_validation(role_dv)
    role_dv.add('E2:E20')

    yn_dv = DataValidation(type="list", formula1='"Yes,No,Partial"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('F2:F20')
    yn_dv.add('G2:G20')
    yn_dv.add('H2:H20')

    status_dv = DataValidation(type="list", formula1='"Approved,Rejected,Pending,Conditional"')
    ws.add_data_validation(status_dv)
    status_dv.add('I2:I20')

    # Format input rows
    for row in range(2, 11):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 15, 15, 25, 20, 20, 18, 18, 15, 15, 18, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_workbook():
    """Generate the complete workbook."""
    logger.info("=" * 70)
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")
    logger.info("=" * 70)

    wb = Workbook()

    # Create all sheets
    create_instructions_sheet(wb.active)
    create_communication_scenarios_sheet(wb.create_sheet())
    create_notification_requirements_sheet(wb.create_sheet())
    create_escalation_matrix_sheet(wb.create_sheet())
    create_approval_workflow_sheet(wb.create_sheet())
    create_communication_templates_sheet(wb.create_sheet())
    create_evidence_register_sheet(wb.create_sheet())
    create_approval_signoff_sheet(wb.create_sheet())

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")
    logger.info("=" * 70)

    return OUTPUT_FILENAME


if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.5-6 External Communications control
# =============================================================================
