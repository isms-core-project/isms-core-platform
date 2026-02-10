#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.5-6.S4 - External Communications Compliance Dashboard
================================================================================

ISO/IEC 27001:2022 Controls A.5.5 & A.5.6: External Communications
Assessment Domain 4 of 4: Compliance Dashboard

This script generates a comprehensive Excel dashboard for monitoring compliance
with external communications requirements including KPIs, metrics, and
audit readiness status per ISO 27001:2022 Controls A.5.5 and A.5.6.
================================================================================
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.5-6.S4"
WORKBOOK_NAME = "External Communications Compliance Dashboard"
CONTROL_ID = "A.5.5-6"
CONTROL_NAME = "Contact with Authorities & Special Interest Groups"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls A.5.5 & A.5.6: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLING CONSTANTS
# =============================================================================
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

METRIC_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
METRIC_FONT = Font(bold=True, size=12, color="FFFFFF")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# WORKBOOK GENERATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws):
    """Create the Instructions sheet."""
    ws.title = "Instructions"

    instructions = [
        ["ISMS-IMP-A.5.5-6.S4 - External Communications Compliance Dashboard"],
        [""],
        ["PURPOSE"],
        ["This dashboard provides a consolidated view of external communications compliance,"],
        ["including KPIs, metrics, and audit readiness status for ISO 27001:2022 Controls A.5.5 and A.5.6."],
        [""],
        ["SHEETS"],
        ["1. Instructions - This guidance sheet"],
        ["2. Executive_Summary - High-level compliance status and key metrics"],
        ["3. Authority_KPIs - Metrics for authority contact management (A.5.5)"],
        ["4. SIG_KPIs - Metrics for special interest group engagement (A.5.6)"],
        ["5. Compliance_Scorecard - Detailed compliance assessment"],
        ["6. Gap_Analysis - Identified gaps and remediation status"],
        ["7. Audit_Readiness - Evidence checklist for audits"],
        ["8. Trend_Analysis - Historical compliance trends"],
        ["9. Evidence_Register - Evidence documentation"],
        ["10. Approval_SignOff - Management approval workflow"],
        [""],
        ["KEY METRICS TRACKED"],
        ["- Authority contact register completeness"],
        ["- Contact verification currency (< 12 months)"],
        ["- SIG membership value assessment"],
        ["- Communication procedure documentation"],
        ["- Notification requirement coverage"],
        ["- Evidence availability for audits"],
        [""],
        ["REFRESH FREQUENCY"],
        ["- Executive Summary: Monthly"],
        ["- KPI Dashboards: Quarterly"],
        ["- Gap Analysis: After each assessment"],
        ["- Audit Readiness: Before scheduled audits"],
        [""],
        [f"Generated: {GENERATED_DATE}"],
        [f"Control Reference: {CONTROL_REF}"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["PURPOSE", "SHEETS", "KEY METRICS TRACKED", "REFRESH FREQUENCY"]:
                cell.font = Font(bold=True, size=11)

    ws.column_dimensions['A'].width = 90


def create_executive_summary_sheet(ws):
    """Create the Executive Summary sheet."""
    ws.title = "Executive_Summary"

    # Title row
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="A.5.5-6 External Communications - Executive Summary")
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = METRIC_HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center")

    # Reporting period
    ws.cell(row=2, column=1, value="Reporting Period:").font = Font(bold=True)
    ws.cell(row=2, column=2, value="").fill = INPUT_FILL

    ws.cell(row=2, column=4, value="Last Updated:").font = Font(bold=True)
    ws.cell(row=2, column=5, value="").fill = INPUT_FILL

    # Overall Status Section
    ws.cell(row=4, column=1, value="OVERALL COMPLIANCE STATUS").font = Font(bold=True, size=12)

    summary_headers = ["Control", "Status", "Score", "Trend", "Key Issues"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    summary_data = [
        ("A.5.5 Contact with Authorities", "", "", "", ""),
        ("A.5.6 Contact with Special Interest Groups", "", "", "", ""),
        ("Combined A.5.5-6 Compliance", "", "", "", ""),
    ]

    row = 6
    for data in summary_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if col > 1:
                cell.fill = INPUT_FILL
        row += 1

    # Data validations for status
    status_dv = DataValidation(type="list", formula1='"Compliant,Partially Compliant,Non-Compliant"')
    ws.add_data_validation(status_dv)
    status_dv.add('B6:B8')

    trend_dv = DataValidation(type="list", formula1='"Improving,Stable,Declining"')
    ws.add_data_validation(trend_dv)
    trend_dv.add('D6:D8')

    # Key Metrics Section
    ws.cell(row=10, column=1, value="KEY METRICS SUMMARY").font = Font(bold=True, size=12)

    metrics_headers = ["Metric", "Target", "Current", "Status", "Notes"]
    for col, header in enumerate(metrics_headers, 1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    metrics = [
        ("Authority contacts documented", "100%", "", "", ""),
        ("Contacts verified < 12 months", ">= 95%", "", "", ""),
        ("SIG memberships active", ">= 5", "", "", ""),
        ("Communication procedures documented", "100%", "", "", ""),
        ("Notification requirements mapped", "100%", "", "", ""),
        ("Escalation paths tested", "Annual", "", "", ""),
    ]

    row = 12
    for metric in metrics:
        for col, value in enumerate(metric, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if col > 2:
                cell.fill = INPUT_FILL
        row += 1

    # Column widths
    widths = [40, 15, 15, 15, 45]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_authority_kpis_sheet(ws):
    """Create the Authority KPIs sheet."""
    ws.title = "Authority_KPIs"

    headers = [
        "KPI_ID", "KPI_Name", "Description", "Target", "Current_Value",
        "Measurement_Method", "Frequency", "Data_Source", "Owner",
        "Status", "Trend", "Comments"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate KPIs
    kpis = [
        ("AUTH-KPI-001", "Register Completeness", "% of required authorities documented",
         "100%", "", "Count documented / Required", "Quarterly",
         "Authority Registry", "CISO", "", "", ""),
        ("AUTH-KPI-002", "Contact Currency", "% of contacts verified in last 12 months",
         ">= 95%", "", "Count verified / Total", "Quarterly",
         "Verification Register", "Security Manager", "", "", ""),
        ("AUTH-KPI-003", "Escalation Path Coverage", "% of scenarios with defined escalation",
         "100%", "", "Scenarios with path / Total", "Annual",
         "Escalation Matrix", "CISO", "", "", ""),
        ("AUTH-KPI-004", "Communication Log Currency", "Communications logged in last 6 months",
         ">= 1 per authority", "", "Count per authority", "Semi-annual",
         "Communication Log", "Security Manager", "", "", ""),
        ("AUTH-KPI-005", "Procedure Test Success", "% of procedures tested successfully",
         ">= 90%", "", "Successful tests / Total", "Annual",
         "Test Records", "CISO", "", "", ""),
    ]

    row = 2
    for kpi in kpis:
        for col, value in enumerate(kpi, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if col in [5, 10, 11, 12]:
                cell.fill = INPUT_FILL
        row += 1

    # Add empty input rows
    for row in range(row, row + 10):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validations
    status_dv = DataValidation(type="list", formula1='"On Target,At Risk,Below Target"')
    ws.add_data_validation(status_dv)
    status_dv.add('J2:J30')

    trend_dv = DataValidation(type="list", formula1='"Improving,Stable,Declining"')
    ws.add_data_validation(trend_dv)
    trend_dv.add('K2:K30')

    freq_dv = DataValidation(type="list", formula1='"Monthly,Quarterly,Semi-annual,Annual"')
    ws.add_data_validation(freq_dv)
    freq_dv.add('G2:G30')

    # Column widths
    widths = [15, 25, 40, 12, 15, 30, 15, 20, 18, 12, 12, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_sig_kpis_sheet(ws):
    """Create the SIG KPIs sheet."""
    ws.title = "SIG_KPIs"

    headers = [
        "KPI_ID", "KPI_Name", "Description", "Target", "Current_Value",
        "Measurement_Method", "Frequency", "Data_Source", "Owner",
        "Status", "Trend", "Comments"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate KPIs
    kpis = [
        ("SIG-KPI-001", "Active Memberships", "Number of active SIG memberships",
         ">= 5", "", "Count active memberships", "Quarterly",
         "Groups Registry", "CISO", "", "", ""),
        ("SIG-KPI-002", "Engagement Frequency", "Engagements per quarter per active membership",
         ">= 1", "", "Engagements / Active memberships", "Quarterly",
         "Engagement Log", "Security Manager", "", "", ""),
        ("SIG-KPI-003", "Intelligence Actionability", "% of received intel acted upon",
         ">= 60%", "", "Actioned / Received", "Quarterly",
         "Intelligence Received", "CISO", "", "", ""),
        ("SIG-KPI-004", "Contribution Rate", "Contributions to SIGs per year",
         ">= 4", "", "Count contributions", "Annual",
         "Contribution Log", "CISO", "", "", ""),
        ("SIG-KPI-005", "Membership ROI", "Assessed value vs cost",
         "Positive", "", "Value rating vs cost", "Annual",
         "Groups Registry", "CISO", "", "", ""),
    ]

    row = 2
    for kpi in kpis:
        for col, value in enumerate(kpi, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if col in [5, 10, 11, 12]:
                cell.fill = INPUT_FILL
        row += 1

    # Add empty input rows
    for row in range(row, row + 10):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validations
    status_dv = DataValidation(type="list", formula1='"On Target,At Risk,Below Target"')
    ws.add_data_validation(status_dv)
    status_dv.add('J2:J30')

    trend_dv = DataValidation(type="list", formula1='"Improving,Stable,Declining"')
    ws.add_data_validation(trend_dv)
    trend_dv.add('K2:K30')

    # Column widths
    widths = [15, 25, 40, 12, 15, 35, 15, 20, 18, 12, 12, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_compliance_scorecard_sheet(ws):
    """Create the Compliance Scorecard sheet."""
    ws.title = "Compliance_Scorecard"

    headers = [
        "Requirement_ID", "Requirement", "Control_Reference", "Weight",
        "Evidence_Available", "Implementation_Status", "Score",
        "Max_Score", "Gap_Description", "Remediation_Status"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate requirements
    requirements = [
        ("REQ-001", "Authority contacts register maintained", "A.5.5", "10", "", "", "", "10", "", ""),
        ("REQ-002", "Contact details verified annually", "A.5.5", "10", "", "", "", "10", "", ""),
        ("REQ-003", "Escalation paths documented", "A.5.5", "8", "", "", "", "8", "", ""),
        ("REQ-004", "Communication procedures defined", "A.5.5", "10", "", "", "", "10", "", ""),
        ("REQ-005", "Regulatory notification requirements mapped", "A.5.5", "10", "", "", "", "10", "", ""),
        ("REQ-006", "SIG membership register maintained", "A.5.6", "8", "", "", "", "8", "", ""),
        ("REQ-007", "SIG engagement logged", "A.5.6", "6", "", "", "", "6", "", ""),
        ("REQ-008", "Intelligence received and actioned", "A.5.6", "8", "", "", "", "8", "", ""),
        ("REQ-009", "Contributions appropriately approved", "A.5.6", "6", "", "", "", "6", "", ""),
        ("REQ-010", "Value assessment conducted", "A.5.6", "6", "", "", "", "6", "", ""),
    ]

    row = 2
    for req in requirements:
        for col, value in enumerate(req, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if col in [5, 6, 7, 9, 10]:
                cell.fill = INPUT_FILL
        row += 1

    # Add empty input rows
    for row in range(row, row + 10):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validations
    evidence_dv = DataValidation(type="list", formula1='"Yes,Partial,No"')
    ws.add_data_validation(evidence_dv)
    evidence_dv.add('E2:E30')

    impl_dv = DataValidation(type="list", formula1='"Fully Implemented,Partially Implemented,Not Implemented,Not Applicable"')
    ws.add_data_validation(impl_dv)
    impl_dv.add('F2:F30')

    remed_dv = DataValidation(type="list", formula1='"Not Required,In Progress,Planned,Completed"')
    ws.add_data_validation(remed_dv)
    remed_dv.add('J2:J30')

    # Column widths
    widths = [12, 40, 15, 8, 18, 22, 8, 10, 40, 18]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_gap_analysis_sheet(ws):
    """Create the Gap Analysis sheet."""
    ws.title = "Gap_Analysis"

    headers = [
        "Gap_ID", "Requirement_Reference", "Gap_Description", "Risk_Level",
        "Root_Cause", "Remediation_Action", "Owner", "Target_Date",
        "Status", "Progress", "Evidence_of_Closure", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    risk_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(risk_dv)
    risk_dv.add('D2:D50')

    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Closed,Deferred"')
    ws.add_data_validation(status_dv)
    status_dv.add('I2:I50')

    progress_dv = DataValidation(type="list", formula1='"0%,25%,50%,75%,100%"')
    ws.add_data_validation(progress_dv)
    progress_dv.add('J2:J50')

    # Format input rows
    for row in range(2, 31):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 20, 45, 12, 35, 45, 18, 15, 12, 10, 30, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_audit_readiness_sheet(ws):
    """Create the Audit Readiness sheet."""
    ws.title = "Audit_Readiness"

    headers = [
        "Check_ID", "Audit_Requirement", "Evidence_Required", "Evidence_Location",
        "Evidence_Available", "Last_Reviewed", "Reviewer", "Status", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate audit checks
    checks = [
        ("AUD-001", "Authority register maintained", "Completed Authority Registry workbook",
         "", "", "", "", "", ""),
        ("AUD-002", "Contact verification evidence", "Verification Register with dates",
         "", "", "", "", "", ""),
        ("AUD-003", "Communication log complete", "Communication Log entries",
         "", "", "", "", "", ""),
        ("AUD-004", "SIG membership documentation", "Membership certificates, contracts",
         "", "", "", "", "", ""),
        ("AUD-005", "Engagement evidence", "Meeting minutes, attendance records",
         "", "", "", "", "", ""),
        ("AUD-006", "Intelligence handling records", "Intelligence Received log",
         "", "", "", "", "", ""),
        ("AUD-007", "Contribution approvals", "Contribution Log with approvals",
         "", "", "", "", "", ""),
        ("AUD-008", "Escalation procedures documented", "Escalation Matrix",
         "", "", "", "", "", ""),
        ("AUD-009", "Notification procedures documented", "Notification Requirements sheet",
         "", "", "", "", "", ""),
        ("AUD-010", "Management approval of registers", "Approval SignOff sheets",
         "", "", "", "", "", ""),
    ]

    row = 2
    for check in checks:
        for col, value in enumerate(check, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if col >= 4:
                cell.fill = INPUT_FILL
        row += 1

    # Add empty input rows
    for row in range(row, row + 10):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Data validations
    avail_dv = DataValidation(type="list", formula1='"Yes,Partial,No"')
    ws.add_data_validation(avail_dv)
    avail_dv.add('E2:E30')

    status_dv = DataValidation(type="list", formula1='"Ready,Action Required,Not Applicable"')
    ws.add_data_validation(status_dv)
    status_dv.add('H2:H30')

    # Column widths
    widths = [12, 35, 40, 35, 18, 15, 18, 18, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_trend_analysis_sheet(ws):
    """Create the Trend Analysis sheet."""
    ws.title = "Trend_Analysis"

    headers = [
        "Period", "Authority_Compliance_Score", "SIG_Compliance_Score",
        "Combined_Score", "Key_Changes", "Notable_Events", "Actions_Taken"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Pre-populate periods
    periods = ["Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026",
               "Q1 2027", "Q2 2027", "Q3 2027", "Q4 2027"]

    row = 2
    for period in periods:
        ws.cell(row=row, column=1, value=period).border = THIN_BORDER
        for col in range(2, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        row += 1

    # Column widths
    widths = [12, 25, 25, 18, 40, 40, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_evidence_register_sheet(ws):
    """Create the Evidence Register sheet."""
    ws.title = "Evidence_Register"

    headers = [
        "Evidence_ID", "Evidence_Type", "Description", "Related_Requirement",
        "Date_Created", "Created_By", "Storage_Location",
        "Retention_Period", "Review_Date", "Status", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    type_dv = DataValidation(
        type="list",
        formula1='"KPI Report,Audit Report,Assessment Record,Approval Record,Trend Analysis,Other"'
    )
    ws.add_data_validation(type_dv)
    type_dv.add('B2:B100')

    status_dv = DataValidation(type="list", formula1='"Current,Archived,Pending Review"')
    ws.add_data_validation(status_dv)
    status_dv.add('J2:J100')

    # Format input rows
    for row in range(2, 31):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 18, 40, 20, 15, 20, 40, 15, 15, 15, 35]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_approval_signoff_sheet(ws):
    """Create the Approval Sign-Off sheet."""
    ws.title = "Approval_SignOff"

    headers = [
        "Approval_ID", "Review_Period", "Review_Date", "Reviewer_Name",
        "Reviewer_Role", "Dashboard_Complete", "KPIs_Current",
        "Gaps_Addressed", "Approval_Status", "Signature_Date",
        "Next_Review_Date", "Comments"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data validations
    role_dv = DataValidation(
        type="list",
        formula1='"CISO,Security Manager,DPO,Compliance Officer,Internal Audit,CEO"'
    )
    ws.add_data_validation(role_dv)
    role_dv.add('E2:E20')

    yn_dv = DataValidation(type="list", formula1='"Yes,No,Partial"')
    ws.add_data_validation(yn_dv)
    yn_dv.add('F2:F20')
    yn_dv.add('G2:G20')
    yn_dv.add('H2:H20')

    status_dv = DataValidation(type="list", formula1='"Approved,Rejected,Pending,Conditional"')
    ws.add_data_validation(status_dv)
    status_dv.add('I2:I20')

    # Format input rows
    for row in range(2, 11):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER

    # Column widths
    widths = [12, 15, 15, 25, 20, 18, 15, 15, 15, 15, 18, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_workbook():
    """Generate the complete workbook."""
    logger.info("=" * 70)
    logger.info(f"Generating {DOCUMENT_ID} - {WORKBOOK_NAME}")
    logger.info("=" * 70)

    wb = Workbook()

    # Create all sheets
    create_instructions_sheet(wb.active)
    create_executive_summary_sheet(wb.create_sheet())
    create_authority_kpis_sheet(wb.create_sheet())
    create_sig_kpis_sheet(wb.create_sheet())
    create_compliance_scorecard_sheet(wb.create_sheet())
    create_gap_analysis_sheet(wb.create_sheet())
    create_audit_readiness_sheet(wb.create_sheet())
    create_trend_analysis_sheet(wb.create_sheet())
    create_evidence_register_sheet(wb.create_sheet())
    create_approval_signoff_sheet(wb.create_sheet())

    # Save workbook
    wb.save(OUTPUT_FILENAME)
    logger.info(f"Workbook saved: {OUTPUT_FILENAME}")
    logger.info("=" * 70)

    return OUTPUT_FILENAME


if __name__ == "__main__":
    generate_workbook()

# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.5-6 External Communications control
# =============================================================================
