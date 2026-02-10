#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.4.S2 - Physical Environmental Protection Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.4: Physical Security Monitoring
Assessment Domain 2 of 4: Environmental Monitoring and Protection Systems

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific physical security infrastructure, monitoring
systems, and assessment requirements.

Key customization areas:
1. Access control systems and reader types (match your actual hardware)
2. CCTV camera models and recording capabilities (adapt to your installations)
3. Intrusion detection sensor types (specific to your alarm systems)
4. Compliance thresholds (aligned with your security requirements)
5. Incident classification criteria (based on your risk profile)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.7.4 Physical Security Monitoring Framework

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (Python standard library)

Usage:
    python3 generate_a74_1_environmental_protection.py

Output:
    ISMS-IMP-A.7.4.S2_Environmental_Protection_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)



# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.4.S2"
WORKBOOK_NAME = "Physical Environmental Protection Assessment"
CONTROL_ID = "A.7.4"
CONTROL_NAME = "Physical Security Monitoring"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
HOURGLASS = '\u23F3'  # ⏳ Hourglass
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# STYLE DEFINITIONS (A.8.24 PATTERN)
# ============================================================================

def setup_styles():
    """Define all cell styles matching A.8.24 pattern."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# ============================================================================
# INSTRUCTIONS SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.7.4.S2 - Physical Environmental Protection Assessment\n"
        "ISO/IEC 27001:2022 - Control A.7.4: Physical Security Monitoring"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.7.4.S2"),
        ("Assessment Area", "Physical Environmental Protection Controls"),
        ("Related Policy", "ISMS-POL-A.7.4"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Quarterly"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "1. Complete each worksheet tab for applicable physical security systems.",
        "2. Use dropdown menus for standardised status entries.",
        "3. Fill in yellow-highlighted cells with your facility information.",
        "4. Document all access control readers, CCTV cameras, and intrusion sensors.",
        "5. Track security incidents and response times.",
        "6. Provide evidence location/path for audit traceability.",
        "7. Summary Dashboard auto-calculates compliance metrics.",
        "8. Obtain final approval in the Approval Sign-Off sheet.",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws[f"A{row+1}"] = "STATUS LEGEND"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    legend = [
        ("Symbol", "Status", "Description"),
        ("{CHECK}", "Compliant", "Fully operational and meets requirements"),
        ("{WARNING}", "Partial", "Functional but requires attention"),
        ("{XMARK}", "Non-Compliant", "Not operational or does not meet requirements"),
        ("N/A", "Not Applicable", "Not required for this facility"),
    ]

    row += 2
    for c, h in enumerate(legend[0], start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for sym, status, desc in legend[1:]:
        ws.cell(row=row, column=1, value=sym).border = styles["border"]
        s = ws.cell(row=row, column=2, value=status)
        d = ws.cell(row=row, column=3, value=desc)
        for cell in (s, d):
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if "Compliant" in status and "Non" not in status:
            s.fill = styles["status_compliant"]["fill"]
        elif "Partial" in status:
            s.fill = styles["status_partial"]["fill"]
        elif "Non-Compliant" in status:
            s.fill = styles["status_noncompliant"]["fill"]
        row += 1

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 50

# ============================================================================
# ACCESS CONTROL SHEET
# ============================================================================

def create_fire_detection_sheet(ws, styles):
    """Create Access Control assessment sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Fire Detection Assessment\nPolicy Requirement: All facility entry points must have electronic access control"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A3:H3")
    ws["A3"] = "Are fire detection systems deployed and tested per schedule in all critical areas?"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    ws["A4"] = "Response:"
    ws["A4"].font = Font(name="Calibri", bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    dv_response = DataValidation(type="list", formula1='"Yes,No,Not Applicable"', allow_blank=False)
    ws.add_data_validation(dv_response)
    dv_response.add(ws["B4"])

    # Column headers
    columns = {
        "Detector ID": 15,
        "Type": 20,
        "Location": 25,
        "Last Tested": 15,
        "Status": 15,
        "Evidence Location": 30,
    }

    row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data rows (100 blank rows for data entry)
    for r in range(7, 107):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    # Status dropdown (column E)
    dv_status = DataValidation(type="list", 
                               formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
                               allow_blank=False)
    ws.add_data_validation(dv_status)
    for r in range(7, 107):
        dv_status.add(ws.cell(row=r, column=5))

    ws.freeze_panes = "A7"

# ============================================================================
# CCTV SHEET
# ============================================================================

def create_water_detection_sheet(ws, styles):
    """Create CCTV Coverage assessment sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Water Leak Detection Assessment\nPolicy Requirement: Video surveillance of all entry/exit points with 90-day retention"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A3:H3")
    ws["A3"] = "Are water leak detection sensors deployed in all critical areas?"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    ws["A4"] = "Response:"
    ws["A4"].font = Font(name="Calibri", bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    dv_response = DataValidation(type="list", formula1='"Yes,No,Not Applicable"', allow_blank=False)
    ws.add_data_validation(dv_response)
    dv_response.add(ws["B4"])

    columns = {
        "Sensor ID": 15,
        "Type": 20,
        "Location": 25,
        "Last Tested": 15,
        "Status": 15,
        "Evidence Location": 30,
    }

    row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(7, 107):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    dv_status = DataValidation(type="list", 
                               formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
                               allow_blank=False)
    ws.add_data_validation(dv_status)
    for r in range(7, 107):
        dv_status.add(ws.cell(row=r, column=6))

    ws.freeze_panes = "A7"

# ============================================================================
# INTRUSION DETECTION SHEET
# ============================================================================

def create_temperature_humidity_sheet(ws, styles):
    """Create Intrusion Detection assessment sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Environmental Monitoring Assessment\nPolicy Requirement: Perimeter and critical area monitoring with 24/7 monitoring"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A3:H3")
    ws["A3"] = "Are temperature and humidity continuously monitored in critical areas?"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    ws["A4"] = "Response:"
    ws["A4"].font = Font(name="Calibri", bold=True)
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    dv_response = DataValidation(type="list", formula1='"Yes,No,Not Applicable"', allow_blank=False)
    ws.add_data_validation(dv_response)
    dv_response.add(ws["B4"])

    columns = {
        "Sensor ID": 15,
        "Location": 25,
        "Temp Range (°C)": 18,
        "Humidity Range (%)": 18,
        "Status": 15,
        "Evidence Location": 30,
    }

    row = 6
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(7, 107):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    dv_status = DataValidation(type="list", 
                               formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant,N/A"',
                               allow_blank=False)
    ws.add_data_validation(dv_status)
    for r in range(7, 107):
        dv_status.add(ws.cell(row=r, column=6))

    ws.freeze_panes = "A7"

# Incidents sheet not applicable for environmental assessment

# Placeholder for future expansion
def create_incidents_sheet_placeholder(ws, styles):
    pass

def create_incidents_sheet_REMOVED(ws, styles):
    """Create Security Incidents tracking sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "Physical Security Incidents\nPolicy Requirement: All incidents logged and investigated with <15min response time"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A3:H3")
    ws["A3"] = "Log all physical security incidents including unauthorised access attempts, tailgating, and alarm triggers."
    ws["A3"].font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    columns = {
        "Incident ID": 15,
        "Date": 12,
        "Type": 20,
        "Severity": 12,
        "Location": 20,
        "Response Time (min)": 18,
        "Resolved": 12,
        "Notes": 35,
    }

    row = 5
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for r in range(6, 106):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    dv_type = DataValidation(type="list", 
                             formula1='"Tailgating,Unauthorised Access,Alarm Trigger,Lost Badge,Forced Entry,Other"',
                             allow_blank=False)
    ws.add_data_validation(dv_type)
    for r in range(6, 106):
        dv_type.add(ws.cell(row=r, column=3))

    dv_severity = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=False)
    ws.add_data_validation(dv_severity)
    for r in range(6, 106):
        dv_severity.add(ws.cell(row=r, column=4))

    ws.freeze_panes = "A6"

# ============================================================================
# SUMMARY DASHBOARD
# ============================================================================

def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with metrics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "ACCESS MONITORING SUMMARY DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Period:"
    ws["A3"].font = Font(name="Calibri", bold=True)
    ws["B3"].fill = styles["input_cell"]["fill"]
    ws["B3"].border = styles["border"]

    ws["A5"] = "Compliance Metrics"
    ws["A5"].font = Font(name="Calibri", size=12, bold=True)

    headers = ["Metric", "Count", f"{CHECK} Compliant", f"{WARNING} Partial", f"{XMARK} Non-Compliant", "N/A", "% Compliant"]
    row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    areas = [
        ("Fire Detection", "Fire Detection", 5),
        ("Water Detection", "Water Detection", 5),
        ("Temperature/Humidity", "Temperature_Humidity", 5),
    ]

    row += 1
    start_data_row = row
    for label, sheet, status_col in areas:
        ws.cell(row=row, column=1, value=label)
        
        status_col_letter = get_column_letter(status_col)
        status_range = f"'{sheet}'!{status_col_letter}7:{status_col_letter}106"

        ws.cell(row=row, column=2, value=f'=COUNTA({status_range})')
        ws.cell(row=row, column=3, value=f'=COUNTIF({status_range},"{CHECK}*")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({status_range},"{WARNING}*")')
        ws.cell(row=row, column=5, value=f'=COUNTIF({status_range},"{XMARK}*")')
        ws.cell(row=row, column=6, value=f'=COUNTIF({status_range},"N/A")')
        ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_data_row}:{get_column_letter(col)}{row-1})")
        cell.font = Font(name="Calibri", bold=True)

    total_pct = ws.cell(row=row, column=7, value=f'=IF((B{row}-F{row})=0,"0%",ROUND(C{row}/(B{row}-F{row})*100,1)&"%")')
    total_pct.font = Font(name="Calibri", bold=True, color="0000FF", size=12)

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 18

    # Incident metrics (manual entry - no separate incidents sheet for environmental assessment)
    row += 3
    ws[f"A{row}"] = "Incident Response Metrics"
    ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)

    row += 1
    ws[f"A{row}"] = "Total Incidents (Last Period):"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws[f"B{row}"] = ""  # Manual entry
    ws[f"B{row}"].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    row += 1
    ws[f"A{row}"] = "Average Response Time (minutes):"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws[f"B{row}"] = ""  # Manual entry
    ws[f"B{row}"].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

# ============================================================================
# EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "List all evidence files/documents referenced in this assessment (audit traceability)."
    ws["A2"].font = Font(name="Calibri", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Evidence ID",
        "Assessment Area",
        "Evidence Type",
        "Description",
        "Location/Path",
        "Date Collected",
        "Collected By",
        "Verification Status",
    ]
    widths = [15, 25, 22, 40, 45, 16, 20, 22]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Configuration file,Screenshot,Access log,Video footage,Test report,Maintenance record,Audit log,Other"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_type)

    dv_ver = DataValidation(
        type="list",
        formula1='"Verified,Pending verification,Not verified,Requires update"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_ver)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"EV-{r-4:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=3))
        dv_ver.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"

# ============================================================================
# APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(ws, styles):
    """Create Approval and Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Summary"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    summary_fields = [
        ("Assessment Document", "ISMS-IMP-A.7.4.S2 - Environmental Protection Assessment"),
        ("Assessment Period", ""),
        ("Overall Compliance Rate", "='Summary Dashboard'!G9"),
        ("Assessment Status", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Approval Workflow"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    approval_headers = ["Role", "Name", "Signature", "Date"]
    row += 3
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    roles = ["Assessor", "Security Manager", "ISO Reviewer", "CISO Approver"]
    row += 1
    for role in roles:
        ws.cell(row=row, column=1, value=role).font = Font(name="Calibri", bold=True)
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15

# ============================================================================
# MAIN WORKBOOK GENERATION
# ============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)
    
    styles = setup_styles()
    
    # Create all sheets
    ws = wb.create_sheet("Instructions & Legend", 0)
    create_instructions_sheet(ws, styles)
    
    ws = wb.create_sheet("Fire Detection")
    create_fire_detection_sheet(ws, styles)
    
    ws = wb.create_sheet("Water Detection")
    create_water_detection_sheet(ws, styles)
    
    ws = wb.create_sheet("Temperature_Humidity")
    create_temperature_humidity_sheet(ws, styles)
    
    # Incidents sheet not needed for environmental assessment
    
    ws = wb.create_sheet("Summary Dashboard")
    create_summary_dashboard(ws, styles)
    
    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)
    
    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_signoff(ws, styles)
    
    return wb

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main entry point for workbook generation."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS Physical Environmental Protection Assessment (A.7.4.2)")
        logger.info("=" * 70)

        wb = create_workbook()
        filename = f"ISMS-IMP-A.7.4.S2_Environmental_Protection_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)

        logger.info("%s SUCCESS: %s", CHECK, filename)
        logger.info("  %s 7 professional worksheets created", BULLET)
        logger.info("  %s A.8.24 styling applied (navy headers, yellow inputs)", BULLET)
        logger.info("  %s 100 data rows per assessment sheet", BULLET)
        logger.info("  %s Automated compliance dashboard with formulas", BULLET)
        logger.info("  %s Data validations and freeze panes configured", BULLET)
        logger.info("  %s Evidence register with audit traceability", BULLET)
        logger.info("  %s 4-level approval workflow", BULLET)
        logger.info("=" * 70)

        return 0

    except Exception as e:
        logger.error("%s ERROR: Failed to generate workbook", XMARK)
        logger.error("Error: %s", str(e))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
