#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.4.S4 - Physical Infrastructure Compliance Dashboard Generator
================================================================================

ISO/IEC 27001:2022 Controls A.7.4/A.7.5/A.7.11: Physical Infrastructure Security
Assessment Domain 4 of 4: Consolidated Compliance Dashboard

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific physical infrastructure assessment requirements,
facility types, and compliance reporting needs.

Key customization areas:
1. Source workbook file naming conventions (match your assessment outputs)
2. Compliance threshold definitions (aligned with your risk appetite)
3. Dashboard KPI calculations (specific to your infrastructure)
4. Evidence consolidation rules (based on your audit requirements)
5. Stakeholder reporting formats (adapted to your governance structure)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.7.4/5/11 Physical Infrastructure Framework

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (Python standard library)

Usage:
    python3 generate_a74_4_compliance_dashboard.py

Output:
    ISMS-IMP-A.7.4.S4_Compliance_Dashboard_YYYYMMDD.xlsx

Prerequisites:
    Completed assessment workbooks must exist:
    - ISMS-IMP-A.7.4.S1_Access_Monitoring.xlsx
    - ISMS-IMP-A.7.4.S2_Environmental_Protection.xlsx
    - ISMS-IMP-A.7.4.S3_Utility_Resilience.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)




# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.4.S4"
WORKBOOK_NAME = "Physical Infrastructure Compliance Dashboard"
CONTROL_ID = "A.7.4"
CONTROL_NAME = "Physical Security Monitoring"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
HOURGLASS = '\u23F3'  # ⏳ Hourglass
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# STYLE DEFINITIONS (CONSISTENT PATTERN)
# ============================================================================

def setup_styles():
    """Define all cell styles matching framework pattern."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# ============================================================================
# INSTRUCTIONS SHEET
# ============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "ISMS-IMP-A.7.4.S4 - Physical Infrastructure Compliance Dashboard\n"
        "ISO/IEC 27001:2022 - Controls A.7.4/A.7.5/A.7.11: Physical Infrastructure Security"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", "ISMS-IMP-A.7.4.S4"),
        ("Dashboard Type", "Physical Infrastructure Compliance Consolidation"),
        ("Related Policies", "ISMS-POL-A.7.4, ISMS-POL-A.7.5, ISMS-POL-A.7.11"),
        ("Version", "1.0"),
        ("Generation Date", datetime.now().strftime("%d.%m.%Y")),
        ("Generated By", ""),
        ("Organisation", ""),
        ("Reporting Period", ""),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        row += 1

    ws["A13"] = "Dashboard Purpose"
    ws["A13"].font = Font(name="Calibri", size=11, bold=True)
    
    purpose_text = (
        f"{BULLET} Consolidates physical security monitoring (A.7.4), environmental protection (A.7.5), "
        f"and utility resilience (A.7.11) assessments into unified executive dashboard\n"
        f"{BULLET} Provides overall physical infrastructure security posture\n"
        f"{BULLET} Identifies critical gaps requiring immediate remediation\n"
        f"{BULLET} Tracks compliance trends across all three control domains\n"
        f"{BULLET} Supports audit evidence collection and management review\n"
    )
    ws["A14"] = purpose_text
    ws["A14"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[14].height = 80

    # CUSTOMIZE: Add organization-specific dashboard usage instructions
    ws["A16"] = "How to Use This Dashboard"
    ws["A16"].font = Font(name="Calibri", size=11, bold=True)
    
    usage_text = (
        f"{ARROW} Executive Dashboard: Overall compliance scores and status indicators\n"
        f"{ARROW} Gap Analysis: Critical findings requiring immediate attention\n"
        f"{ARROW} Risk Register: Physical infrastructure security risks\n"
        f"{ARROW} KPIs & Metrics: Performance indicators across all three controls\n"
        f"{ARROW} Remediation Roadmap: Action items with priorities and timelines\n"
        f"{ARROW} Evidence Register: Consolidated audit evidence from all assessments\n"
    )
    ws["A17"] = usage_text
    ws["A17"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[17].height = 90

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

# ============================================================================
# EXECUTIVE DASHBOARD SHEET
# ============================================================================

def create_executive_dashboard(ws, styles):
    """Create Executive Dashboard with consolidated compliance view."""
    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "Physical Infrastructure Security - Executive Dashboard"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Subheader
    ws.merge_cells("A2:H2")
    ws["A2"] = f"ISO/IEC 27001:2022 Controls A.7.4/A.7.5/A.7.11 {BULLET} Assessment Period: [Specify Period]"
    ws["A2"].font = styles["subheader"]["font"]
    ws["A2"].fill = styles["subheader"]["fill"]
    ws["A2"].alignment = styles["subheader"]["alignment"]

    # Overall Compliance Score Section
    ws["A4"] = "Overall Physical Infrastructure Compliance"
    ws["A4"].font = Font(name="Calibri", size=12, bold=True)
    
    headers = ["Control", "Compliance Score", "Status", "Critical Gaps", "High Priority", "Medium Priority"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # CUSTOMIZE: Link these formulas to actual source workbooks
    control_data = [
        ("A.7.4 - Physical Security Monitoring", "95%", f"{WARNING} Partial", "0", "2", "3"),
        ("A.7.5 - Environmental Protection", "88%", f"{WARNING} Partial", "1", "3", "2"),
        ("A.7.11 - Utility Resilience", "92%", f"{CHECK} Compliant", "0", "1", "4"),
        ("", "", "", "", "", ""),
        ("OVERALL PHYSICAL INFRASTRUCTURE", "92%", f"{CHECK} Compliant", "1", "6", "9"),
    ]

    row = 6
    for data in control_data:
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply conditional formatting based on status
            if "Compliant" in str(value) and CHECK in str(value):
                cell.fill = styles["status_compliant"]["fill"]
            elif "Partial" in str(value):
                cell.fill = styles["status_partial"]["fill"]
            elif "Non-Compliant" in str(value):
                cell.fill = styles["status_noncompliant"]["fill"]
                
            # Bold the overall row
            if row == 10:
                cell.font = Font(bold=True)
        row += 1

    # Key Metrics Section
    ws["A12"] = "Key Physical Infrastructure Metrics"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    metric_headers = ["Metric Category", "Current Value", "Target", "Status", "Trend"]
    for col, header in enumerate(metric_headers, start=1):
        cell = ws.cell(row=13, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # CUSTOMIZE: Link these to actual calculated metrics
    metrics = [
        ("Access Control Coverage", "100%", "100%", f"{CHECK} On Target", f"{ARROW} Stable"),
        ("CCTV System Coverage", "98%", "100%", f"{WARNING} Below Target", f"{ARROW} Improving"),
        ("Intrusion Detection Coverage", "95%", "100%", f"{WARNING} Below Target", f"{ARROW} Stable"),
        ("Fire Suppression Readiness", "100%", "100%", f"{CHECK} On Target", f"{ARROW} Stable"),
        ("Environmental Monitoring", "92%", "95%", f"{WARNING} Below Target", f"{ARROW} Improving"),
        ("Power Uptime (UPS)", "99.99%", "99.99%", f"{CHECK} On Target", f"{ARROW} Stable"),
        ("HVAC Availability", "99.95%", "99.95%", f"{CHECK} On Target", f"{ARROW} Stable"),
        ("Generator Test Success Rate", "100%", "100%", f"{CHECK} On Target", f"{ARROW} Stable"),
    ]

    row = 14
    for metric in metrics:
        for col, value in enumerate(metric, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if CHECK in str(value):
                cell.fill = styles["status_compliant"]["fill"]
            elif WARNING in str(value):
                cell.fill = styles["status_partial"]["fill"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

# ============================================================================
# GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet showing critical findings."""
    # Header
    ws.merge_cells("A1:I1")
    ws["A1"] = "Physical Infrastructure Security - Gap Analysis"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Column headers
    headers = [
        "Gap ID", "Control", "Finding", "Severity", "Impact", 
        "Current State", "Required State", "Remediation Action", "Target Date"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # CUSTOMIZE: Populate with actual gaps from source assessments
    gaps = [
        ("GAP-001", "A.7.5", "Datacenter lacks redundant fire suppression", 
         f"{XMARK} Critical", "High", "Single system", "N+1 redundancy", 
         "Install secondary suppression system", "Q2 2025"),
        ("GAP-002", "A.7.4", "CCTV coverage gap in loading dock area",
         f"{WARNING} High", "Medium", "85% coverage", "100% coverage",
         "Install 3 additional cameras", "Q1 2025"),
        ("GAP-003", "A.7.11", "UPS battery age exceeds 5 years",
         f"{WARNING} High", "Medium", "6 years old", "< 5 years",
         "Replace UPS batteries", "Q1 2025"),
    ]

    row = 4
    for gap in gaps:
        for col, value in enumerate(gap, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            if XMARK in str(value):
                cell.fill = styles["status_noncompliant"]["fill"]
            elif WARNING in str(value):
                cell.fill = styles["status_partial"]["fill"]
        row += 1

    # Column widths
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["H"].width = 35

# ============================================================================
# KPI METRICS SHEET
# ============================================================================

def create_kpi_metrics(ws, styles):
    """Create KPI & Metrics tracking sheet."""
    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "Physical Infrastructure Security - KPIs & Metrics"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Column headers
    headers = ["KPI", "Measurement", "Current", "Target", "Status", "Trend", "Notes"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # CUSTOMIZE: Define KPIs specific to your infrastructure
    kpis = [
        ("Access Control System Uptime", "Percentage", "99.98%", "99.95%", f"{CHECK} Met", f"{ARROW} Stable", ""),
        ("Failed Access Attempts", "Count/Month", "12", "<20", f"{CHECK} Met", f"{ARROW} Decreasing", ""),
        ("Visitor Policy Compliance", "Percentage", "98%", "100%", f"{WARNING} Near", f"{ARROW} Improving", ""),
        ("CCTV System Availability", "Percentage", "99.9%", "99.9%", f"{CHECK} Met", f"{ARROW} Stable", ""),
        ("Video Retention Compliance", "Percentage", "100%", "100%", f"{CHECK} Met", f"{ARROW} Stable", ""),
        ("Environmental Alarms", "Count/Month", "3", "<5", f"{CHECK} Met", f"{ARROW} Stable", ""),
        ("Fire Drill Completion", "Annual Count", "4", "4", f"{CHECK} Met", f"{ARROW} On Track", ""),
        ("Power Uptime", "Percentage", "99.99%", "99.99%", f"{CHECK} Met", f"{ARROW} Stable", ""),
        ("Generator Test Success", "Percentage", "100%", "100%", f"{CHECK} Met", f"{ARROW} Stable", ""),
        ("HVAC Uptime", "Percentage", "99.95%", "99.95%", f"{CHECK} Met", f"{ARROW} Stable", ""),
    ]

    row = 4
    for kpi in kpis:
        for col, value in enumerate(kpi, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left" if col in [1, 7] else "center", 
                                     vertical="center", wrap_text=True)
            
            if CHECK in str(value):
                cell.fill = styles["status_compliant"]["fill"]
            elif WARNING in str(value):
                cell.fill = styles["status_partial"]["fill"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 30

# ============================================================================
# EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register(ws, styles):
    """Create Evidence Register consolidating all physical infrastructure evidence."""
    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "Physical Infrastructure Security - Evidence Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Column headers
    headers = [
        "Evidence ID", "Control", "Evidence Type", "Description", 
        "Collection Date", "Collector", "Location/Link", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    # CUSTOMIZE: Consolidate from source assessment evidence registers
    evidence_samples = [
        ("EV-A74-001", "A.7.4", "Access Logs", "Badge system access logs - Q4 2024", 
         "15.01.2025", "", "/evidence/access/", f"{CHECK} Collected"),
        ("EV-A74-002", "A.7.4", "CCTV Footage", "Security camera audit - Main entrance",
         "15.01.2025", "", "/evidence/cctv/", f"{CHECK} Collected"),
        ("EV-A75-001", "A.7.5", "Inspection Report", "Fire suppression system inspection",
         "10.01.2025", "", "/evidence/fire/", f"{CHECK} Collected"),
        ("EV-A711-001", "A.7.11", "Test Results", "UPS load test results - December 2024",
         "05.01.2025", "", "/evidence/ups/", f"{CHECK} Collected"),
        ("EV-A711-002", "A.7.11", "Test Results", "Generator monthly test - December 2024",
         "05.01.2025", "", "/evidence/generator/", f"{CHECK} Collected"),
    ]

    row = 4
    for evidence in evidence_samples:
        for col, value in enumerate(evidence, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = styles["border"]
            if col == 6:  # Collector field - input cell
                cell.fill = styles["input_cell"]["fill"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["H"].width = 15

# ============================================================================
# APPROVAL SHEET
# ============================================================================

def create_approval_sheet(ws, styles):
    """Create Approval and Sign-Off sheet."""
    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = "Physical Infrastructure Security Dashboard - Approval & Sign-Off"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Approval workflow
    ws["A3"] = "Review and Approval Workflow"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    approval_headers = ["Role", "Name", "Review Date", "Approval Status", "Signature/Comments"]
    for col, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]

    roles = [
        "Physical Security Manager",
        "Facilities Manager", 
        "IT Infrastructure Manager",
        "CISO / Information Security Manager",
        "Risk & Compliance Officer"
    ]

    row = 5
    for role in roles:
        ws.cell(row=row, column=1, value=role).border = styles["border"]
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 40

# ============================================================================
# MAIN WORKBOOK GENERATOR
# ============================================================================

def create_dashboard():
    """Generate complete physical infrastructure compliance dashboard."""
    logger.info(f"\n{HOURGLASS} Generating Physical Infrastructure Compliance Dashboard...")
    
    wb = Workbook()
    styles = setup_styles()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create all sheets
    sheets = [
        ("Instructions & Legend", create_instructions_sheet),
        ("Executive Dashboard", create_executive_dashboard),
        ("Gap Analysis", create_gap_analysis),
        ("KPIs & Metrics", create_kpi_metrics),
        ("Evidence Register", create_evidence_register),
        ("Approval & Sign-Off", create_approval_sheet),
    ]
    
    for sheet_name, create_func in sheets:
        ws = wb.create_sheet(sheet_name)
        create_func(ws, styles)
        logger.info(f"  {CHECK} Created sheet: {sheet_name}")
    
    return wb

# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == "__main__":
    try:
        logger.info("=" * 80)
        logger.info("ISMS Physical Infrastructure Compliance Dashboard Generator")
        logger.info("ISO/IEC 27001:2022 - Controls A.7.4/A.7.5/A.7.11")
        logger.info("=" * 80)
        
        wb = create_dashboard()
        
        # Generate filename with date
        date_suffix = datetime.now().strftime("%Y%m%d")
        filename = f"ISMS-IMP-A.7.4.S4_Compliance_Dashboard_{date_suffix}.xlsx"
        
        wb.save(filename)
        logger.info(f"\n{CHECK} SUCCESS: Dashboard generated successfully")
        logger.info(f"  File: {filename}")
        logger.info(f"\n{ARROW} Next steps:")
        logger.info(f"  1. Review Executive Dashboard for overall compliance status")
        logger.info(f"  2. Address critical gaps identified in Gap Analysis")
        logger.info(f"  3. Update KPIs with actual measurements from source assessments")
        logger.info(f"  4. Complete Evidence Register with all collected evidence")
        logger.info(f"  5. Obtain stakeholder approvals in Approval & Sign-Off sheet")
        
    except Exception as e:
        logger.error(f"\n{XMARK} ERROR: Dashboard generation failed")
        logger.info(f"  Details: {e}")
        sys.exit(1)

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
