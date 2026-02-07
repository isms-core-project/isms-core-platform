#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.14.S3 - Secure Equipment Disposal Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.14: Secure Disposal or Re-Use of Equipment
Assessment Domain 3 of 4: Data Sanitisation and Equipment Destruction

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates an assessment workbook for documenting disposal requirements,
sanitisation tools, service providers, and equipment disposal tracking.

Reference: ISMS-IMP-A.7.6-7-14-S3 Specification

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    pip3 install openpyxl

Usage:
    python3 generate_a76_3_equipment_disposal.py

Output:
    ISMS-IMP-A.7.14.S3_Equipment_Disposal_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.14.S3"
WORKBOOK_NAME = "Equipment Disposal Assessment"
CONTROL_ID = "A.7.14"
CONTROL_NAME = "Secure Disposal or Re-Use of Equipment"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = '\u2705'
XMARK = '\u274C'
WARNING = '\u26A0'
BULLET = '\u2022'

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions & Legend sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Secure Equipment Disposal and Re-Use"),
        ("Related Policy", "ISMS-POL-A.7.6-7-14"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
        ("Review Cycle", "Semi-Annual"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+1}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row+1}"].font = Font(name="Calibri", size=12, bold=True)

    instructions = [
        "1. Document disposal requirements by equipment type in Sheet 2.",
        "2. List approved sanitisation tools in Sheet 3.",
        "3. Document disposal service providers in Sheet 4.",
        "4. Log all equipment disposals in Sheet 5 (Disposal Log).",
        "5. Review Summary Dashboard (Sheet 6) for compliance metrics.",
        "6. Document evidence (certificates) in Sheet 7.",
        "7. Obtain approvals in Sheet 8.",
    ]

    row += 2
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40


def create_disposal_requirements(ws, styles):
    """Create Disposal Requirements Matrix sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Disposal Requirements Matrix\nDisposal methods by equipment type and data classification"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Equipment Type": 22,
        "Storage Type": 18,
        "CONFIDENTIAL Method": 25,
        "INTERNAL Method": 25,
        "PUBLIC Method": 20,
        "Verification Required": 18,
        "Certificate Required": 18,
        "Implementation Status": 20,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate equipment types
    equipment = [
        ("Hard Drives (HDD)", "HDD", "Physical Destruction", "3-Pass Overwrite or Destruction", "Format"),
        ("Solid State Drives (SSD)", "SSD", "Physical Destruction", "Crypto Erase or Destruction", "Secure Erase"),
        ("Mobile Devices", "Flash", "Physical Destruction", "Factory Reset + Verification", "Factory Reset"),
        ("USB/Removable Media", "Flash", "Physical Destruction", "Secure Overwrite or Destruction", "Format"),
        ("Printers/Copiers", "HDD", "HDD Removal + Destruction", "HDD Removal + Secure Wipe", "Clear Memory"),
        ("Network Equipment", "Config", "N/A", "Config Wipe + Verification", "Config Reset"),
        ("Servers", "HDD/SSD", "Physical Destruction", "Secure Wipe or Destruction", "Format + Reinstall"),
        ("Laptops", "HDD/SSD", "Physical Destruction", "Secure Wipe or Destruction", "Factory Reset"),
        ("Desktops", "HDD/SSD", "Physical Destruction", "Secure Wipe or Destruction", "Factory Reset"),
    ]

    dv_storage = DataValidation(type="list", formula1='"HDD,SSD,Flash,Memory,Config,Mixed"', allow_blank=True)
    dv_method = DataValidation(type="list", formula1='"Physical Destruction,Crypto Erase,Secure Overwrite,3-Pass Overwrite,Factory Reset,Format,Config Wipe,N/A"', allow_blank=True)
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_impl = DataValidation(type="list", formula1='"Implemented,Partial,Not Implemented"', allow_blank=True)
    ws.add_data_validation(dv_storage)
    ws.add_data_validation(dv_method)
    ws.add_data_validation(dv_yesno)
    ws.add_data_validation(dv_impl)

    row = 4
    for eq_type, storage, conf, internal, public in equipment:
        ws.cell(row=row, column=1, value=eq_type).border = styles["border"]
        ws.cell(row=row, column=2, value=storage).border = styles["border"]
        ws.cell(row=row, column=3, value=conf).border = styles["border"]
        ws.cell(row=row, column=4, value=internal).border = styles["border"]
        ws.cell(row=row, column=5, value=public).border = styles["border"]
        for c in range(6, 10):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_yesno.add(ws.cell(row=row, column=6))
        dv_yesno.add(ws.cell(row=row, column=7))
        dv_impl.add(ws.cell(row=row, column=8))
        row += 1

    # Additional blank rows
    for r in range(row, row + 41):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_storage.add(ws.cell(row=r, column=2))
        dv_method.add(ws.cell(row=r, column=3))
        dv_method.add(ws.cell(row=r, column=4))
        dv_method.add(ws.cell(row=r, column=5))
        dv_yesno.add(ws.cell(row=r, column=6))
        dv_yesno.add(ws.cell(row=r, column=7))
        dv_impl.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"


def create_disposal_tools(ws, styles):
    """Create Disposal Tools sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "Approved Disposal Tools and Methods\nData sanitisation tools and physical destruction methods"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Tool/Method Name": 30,
        "Tool Type": 22,
        "Applicable Equipment": 25,
        "Standard/Method": 25,
        "Verification Method": 30,
        "Approved Version": 15,
        "Last Tested": 15,
        "Compliant": 12,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate common tools
    tools = [
        ("DBAN (Darik's Boot and Nuke)", "Software - Overwrite", "HDD", "DoD 5220.22-M (3-pass)", "Software verification report"),
        ("Blancco Drive Eraser", "Software - Overwrite", "HDD, SSD", "NIST 800-88 Purge", "Certificate generation"),
        ("BitRaser", "Software - Crypto Erase", "SSD, Flash", "NIST 800-88 Purge", "Certificate generation"),
        ("Physical Shredder", "Hardware - Shredder", "All Storage Media", "Physical Destruction", "Visual inspection + certificate"),
        ("Degausser", "Hardware - Degausser", "Magnetic Media (HDD, Tape)", "Magnetic Erasure", "Visual inspection"),
    ]

    dv_type = DataValidation(type="list", formula1='"Software - Overwrite,Software - Crypto Erase,Hardware - Degausser,Hardware - Shredder,Manual Process"', allow_blank=True)
    dv_compliant = DataValidation(type="list", formula1='"Yes,Partial,No"', allow_blank=True)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_compliant)

    row = 4
    for tool, ttype, equip, standard, verify in tools:
        ws.cell(row=row, column=1, value=tool).border = styles["border"]
        ws.cell(row=row, column=2, value=ttype).border = styles["border"]
        ws.cell(row=row, column=3, value=equip).border = styles["border"]
        ws.cell(row=row, column=4, value=standard).border = styles["border"]
        ws.cell(row=row, column=5, value=verify).border = styles["border"]
        for c in range(6, 10):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_compliant.add(ws.cell(row=row, column=8))
        row += 1

    # Additional blank rows
    for r in range(row, row + 45):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_type.add(ws.cell(row=r, column=2))
        dv_compliant.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"


def create_service_providers(ws, styles):
    """Create Service Providers sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "Disposal Service Providers\nVendor management for equipment destruction services"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Provider Name": 30,
        "Service Type": 22,
        "Contract Status": 15,
        "Contract Expiry": 15,
        "Certificate Provided": 20,
        "On-Site Option": 15,
        "Chain of Custody": 18,
        "Last Audit/Review": 15,
        "Status": 15,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_service = DataValidation(type="list", formula1='"On-site Destruction,Off-site Destruction,Recycling with Destruction,Certificate Only"', allow_blank=True)
    dv_contract = DataValidation(type="list", formula1='"Active,Expired,Under Review,Pending"', allow_blank=True)
    dv_cert = DataValidation(type="list", formula1='"Yes - Per Item,Yes - Per Batch,No"', allow_blank=True)
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_custody = DataValidation(type="list", formula1='"Documented,Partial,Not Documented"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1=f'"{CHECK} Compliant,{WARNING} Partial,{XMARK} Non-Compliant"', allow_blank=True)
    ws.add_data_validation(dv_service)
    ws.add_data_validation(dv_contract)
    ws.add_data_validation(dv_cert)
    ws.add_data_validation(dv_yesno)
    ws.add_data_validation(dv_custody)
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_service.add(ws.cell(row=r, column=2))
        dv_contract.add(ws.cell(row=r, column=3))
        dv_cert.add(ws.cell(row=r, column=5))
        dv_yesno.add(ws.cell(row=r, column=6))
        dv_custody.add(ws.cell(row=r, column=7))
        dv_status.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "A4"


def create_disposal_log(ws, styles):
    """Create Disposal Log sheet."""
    ws.merge_cells("A1:O1")
    ws["A1"] = "Equipment Disposal Log\nTrack all equipment disposals with certificates and verification"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = {
        "Disposal ID": 18,
        "Date": 12,
        "Asset Tag": 15,
        "Equipment Type": 15,
        "Make/Model": 25,
        "Serial Number": 20,
        "Data Classification": 18,
        "Disposal Method": 22,
        "Destination": 18,
        "Certificate Obtained": 18,
        "Certificate Ref": 20,
        "Verified By": 18,
        "Asset Updated": 15,
        "Compliant": 12,
        "Notes": 35,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_equip = DataValidation(type="list", formula1='"Laptop,Desktop,Server,Mobile,Printer,Network,Storage,USB/Media,Other"', allow_blank=True)
    dv_class = DataValidation(type="list", formula1='"CONFIDENTIAL,INTERNAL,PUBLIC,Unknown"', allow_blank=True)
    dv_method = DataValidation(type="list", formula1='"Physical Destruction,Secure Overwrite,Crypto Erase,Factory Reset,Format,Config Wipe"', allow_blank=True)
    dv_dest = DataValidation(type="list", formula1='"Vendor Destruction,Internal Re-Use,Donation,Sale,Recycling"', allow_blank=True)
    dv_cert = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=True)
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_equip)
    ws.add_data_validation(dv_class)
    ws.add_data_validation(dv_method)
    ws.add_data_validation(dv_dest)
    ws.add_data_validation(dv_cert)
    ws.add_data_validation(dv_yesno)

    for r in range(4, 204):  # 200 rows for disposal log
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_equip.add(ws.cell(row=r, column=4))
        dv_class.add(ws.cell(row=r, column=7))
        dv_method.add(ws.cell(row=r, column=8))
        dv_dest.add(ws.cell(row=r, column=9))
        dv_cert.add(ws.cell(row=r, column=10))
        dv_yesno.add(ws.cell(row=r, column=13))
        # Compliance formula - simplified check
        ws.cell(row=r, column=14, value=f'=IF(OR(G{r}="",H{r}=""),"",IF(AND(G{r}="CONFIDENTIAL",H{r}="Physical Destruction",J{r}="Yes"),"Yes",IF(G{r}="PUBLIC","Yes",IF(J{r}="N/A","Yes","Check"))))')

    ws.freeze_panes = "A4"


def create_summary_dashboard(ws, styles):
    """Create Summary Dashboard with metrics."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "EQUIPMENT DISPOSAL - SUMMARY DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Period:"
    ws["A3"].font = Font(name="Calibri", bold=True)
    ws["B3"].fill = styles["input_cell"]["fill"]
    ws["B3"].border = styles["border"]

    ws["A5"] = "Disposal Metrics"
    ws["A5"].font = Font(name="Calibri", size=12, bold=True)

    ws["A7"] = "Total Equipment Disposed:"
    ws["B7"] = "=COUNTA('Disposal Log'!A4:A203)"

    ws["A8"] = "By Data Classification:"
    ws["A9"] = "  - CONFIDENTIAL:"
    ws["B9"] = '=COUNTIF(\'Disposal Log\'!G4:G203,"CONFIDENTIAL")'
    ws["A10"] = "  - INTERNAL:"
    ws["B10"] = '=COUNTIF(\'Disposal Log\'!G4:G203,"INTERNAL")'
    ws["A11"] = "  - PUBLIC:"
    ws["B11"] = '=COUNTIF(\'Disposal Log\'!G4:G203,"PUBLIC")'
    ws["A12"] = "  - Unknown:"
    ws["B12"] = '=COUNTIF(\'Disposal Log\'!G4:G203,"Unknown")'

    ws["A14"] = "Certificate Collection"
    ws["A14"].font = Font(name="Calibri", size=12, bold=True)

    ws["A15"] = "Certificates Obtained:"
    ws["B15"] = '=COUNTIF(\'Disposal Log\'!J4:J203,"Yes")'
    ws["A16"] = "Certificates Required (CONFIDENTIAL):"
    ws["B16"] = '=COUNTIF(\'Disposal Log\'!G4:G203,"CONFIDENTIAL")'
    ws["A17"] = "Certificate Collection Rate:"
    ws["B17"] = '=IF(B16=0,"N/A",ROUND(B15/B16*100,1)&"%")'
    ws["B17"].font = Font(name="Calibri", bold=True, color="0000FF")

    ws["A19"] = "Service Provider Status"
    ws["A19"].font = Font(name="Calibri", size=12, bold=True)

    ws["A20"] = "Active Providers:"
    ws["B20"] = '=COUNTIF(\'Service Providers\'!C4:C53,"Active")'
    ws["A21"] = "Expired Contracts:"
    ws["B21"] = '=COUNTIF(\'Service Providers\'!C4:C53,"Expired")'

    for col in ["A", "B"]:
        ws.column_dimensions[col].width = 35


def create_evidence_register(ws, styles):
    """Create Evidence Register sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "List all evidence including destruction certificates."
    ws["A2"].font = Font(name="Calibri", italic=True)

    headers = [
        "Evidence ID", "Evidence Type", "Description", "Related Disposal ID",
        "File Name", "File Location", "Collection Date", "Collected By",
        "Retention Period", "Notes",
    ]
    widths = [15, 22, 40, 20, 30, 45, 16, 20, 18, 35]

    row = 4
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Destruction Certificate,Sanitisation Report,Asset Record,Vendor Contract,Chain of Custody,Photo,Screenshot,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)

    for r in range(5, 205):  # More rows for certificates
        ws.cell(row=r, column=1, value=f"EV-{r-4:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        dv_type.add(ws.cell(row=r, column=2))

    ws.freeze_panes = "A5"


def create_approval_signoff(ws, styles):
    """Create Approval and Sign-Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Assessment Summary"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Equipment Disposed", "='Summary Dashboard'!B7"),
        ("Certificate Collection Rate", "='Summary Dashboard'!B17"),
        ("Assessment Status", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Approval Workflow"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    approval_headers = ["Role", "Name", "Signature", "Date"]
    row += 3
    for col_idx, header in enumerate(approval_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    roles = ["Assessor", "IT Operations Manager", "CISO", "Compliance Officer"]
    row += 1
    for role in roles:
        ws.cell(row=row, column=1, value=role).font = Font(name="Calibri", bold=True)
        for col in range(2, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15


# =============================================================================
# MAIN WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    styles = setup_styles()

    ws = wb.create_sheet("Instructions & Legend", 0)
    create_instructions_sheet(ws, styles)

    ws = wb.create_sheet("Disposal Requirements")
    create_disposal_requirements(ws, styles)

    ws = wb.create_sheet("Disposal Tools")
    create_disposal_tools(ws, styles)

    ws = wb.create_sheet("Service Providers")
    create_service_providers(ws, styles)

    ws = wb.create_sheet("Disposal Log")
    create_disposal_log(ws, styles)

    ws = wb.create_sheet("Summary Dashboard")
    create_summary_dashboard(ws, styles)

    ws = wb.create_sheet("Evidence Register")
    create_evidence_register(ws, styles)

    ws = wb.create_sheet("Approval Sign-Off")
    create_approval_signoff(ws, styles)

    return wb


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main entry point for workbook generation."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS Equipment Disposal Assessment (A.7.14.S3)")
        logger.info("=" * 70)

        wb = create_workbook()
        wb.save(OUTPUT_FILENAME)

        logger.info("%s SUCCESS: %s", CHECK, OUTPUT_FILENAME)
        logger.info("  %s 8 professional worksheets created", BULLET)
        logger.info("  %s Disposal requirements matrix by classification", BULLET)
        logger.info("  %s Approved tools with standards reference", BULLET)
        logger.info("  %s Service provider management", BULLET)
        logger.info("  %s 200-row disposal log with certificate tracking", BULLET)
        logger.info("  %s Automated summary dashboard", BULLET)
        logger.info("=" * 70)

        return 0

    except Exception as e:
        logger.error("%s ERROR: Failed to generate workbook", XMARK)
        logger.error("Error: %s", str(e))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation with constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
