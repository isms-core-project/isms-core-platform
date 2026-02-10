#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.12-13.S4 - Infrastructure Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.7.12 & A.7.13: Cabling Security and Equipment Maintenance
Assessment Domain 4 of 4: Consolidated Compliance Dashboard

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script generates the Infrastructure Compliance Dashboard workbook that
consolidates compliance data from S1 (Cabling Security), S2 (Equipment Maintenance),
and S3 (Maintenance Schedule) assessments.

Reference Pattern: Based on ISMS-A.7.4-5-11 Physical Infrastructure Framework

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    pip3 install openpyxl

Usage:
    python3 generate_a712_4_compliance_dashboard.py

Output:
    ISMS-IMP-A.7.12-13.S4_Compliance_Dashboard_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.12-13.S4"
WORKBOOK_NAME = "Infrastructure Compliance Dashboard"
CONTROL_ID = "A.7.12-13"
CONTROL_NAME = "Cabling Security and Equipment Maintenance"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = '\u2705'
XMARK = '\u274C'
WARNING = '\u26A0'
BULLET = '\u2022'

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "score_excellent": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "score_good": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "score_acceptable": {"fill": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")},
        "score_poor": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "border": border_thin,
    }
    return styles

# =============================================================================
# EXECUTIVE SUMMARY SHEET
# =============================================================================

def create_executive_summary_sheet(ws, styles):
    """Create Executive Summary sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = (
        f"A.7.12-13 INFRASTRUCTURE COMPLIANCE DASHBOARD\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 50

    # Assessment period
    ws["A3"] = "Assessment Period:"
    ws["A3"].font = Font(name="Calibri", bold=True)
    ws["B3"].fill = styles["input_cell"]["fill"]
    ws["B3"].border = styles["border"]

    ws["D3"] = "Last Updated:"
    ws["D3"].font = Font(name="Calibri", bold=True)
    ws["E3"].fill = styles["input_cell"]["fill"]
    ws["E3"].border = styles["border"]

    # Overall compliance section
    ws["A5"] = "Overall Compliance"
    ws["A5"].font = Font(name="Calibri", size=14, bold=True)

    ws["A7"] = "Overall Infrastructure Compliance Score:"
    ws["A7"].font = Font(name="Calibri", size=12, bold=True)

    ws["D7"] = ""  # Will be manually entered or calculated
    ws["D7"].font = Font(name="Calibri", size=24, bold=True)
    ws["D7"].fill = styles["input_cell"]["fill"]
    ws["D7"].border = styles["border"]
    ws["D7"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A9"] = "Status:"
    ws["A9"].font = Font(name="Calibri", bold=True)
    ws["B9"] = ""  # Excellent/Good/Acceptable/Non-Compliant
    ws["B9"].fill = styles["input_cell"]["fill"]
    ws["B9"].border = styles["border"]

    # Domain breakdown
    ws["A11"] = "Domain Breakdown"
    ws["A11"].font = Font(name="Calibri", size=12, bold=True)

    domains = [
        ("Cabling Security (A.7.12)", "40%", ""),
        ("Equipment Maintenance (A.7.13)", "60%", ""),
    ]

    row = 12
    headers = ["Domain", "Weight", "Score", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for domain, weight, score in domains:
        ws.cell(row=row, column=1, value=domain).border = styles["border"]
        ws.cell(row=row, column=2, value=weight).border = styles["border"]
        ws.cell(row=row, column=3, value=score)
        ws.cell(row=row, column=3).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=3).border = styles["border"]
        ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=4).border = styles["border"]
        row += 1

    # Key risk indicators
    ws[f"A{row+2}"] = "Key Risk Indicators"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    row += 3
    indicators = [
        ("Critical Equipment Overdue (>14 days)", ""),
        ("Cabling Infrastructure Unprotected", ""),
        ("Wiring Closets Without Access Control", ""),
        ("Remote Access Without Logging", ""),
    ]

    for indicator, value in indicators:
        ws.cell(row=row, column=1, value=indicator).border = styles["border"]
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]
        row += 1

    # Gap summary
    ws[f"A{row+2}"] = "Gap Summary"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    row += 3
    gap_headers = ["Priority", "Count"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    priorities = ["Critical", "High", "Medium", "Low"]
    for priority in priorities:
        ws.cell(row=row, column=1, value=priority).border = styles["border"]
        ws.cell(row=row, column=2, value="")
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20

# =============================================================================
# CABLING SECURITY SHEET
# =============================================================================

def create_cabling_security_sheet(ws, styles):
    """Create Cabling Security detail sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "A.7.12 Cabling Security Compliance"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A3"] = "Score Summary"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    row = 4
    headers = ["Domain", "Weight", "Score", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    domains = [
        ("Cable Pathways", "30%"),
        ("Physical Protection", "25%"),
        ("Access Controls", "25%"),
        ("Documentation", "20%"),
        ("DOMAIN TOTAL", "100%"),
    ]

    row += 1
    for domain, weight in domains:
        ws.cell(row=row, column=1, value=domain)
        if domain == "DOMAIN TOTAL":
            ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True)
        ws.cell(row=row, column=1).border = styles["border"]

        ws.cell(row=row, column=2, value=weight).border = styles["border"]

        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=3).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=3).border = styles["border"]

        ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=4).border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Detailed Metrics"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    row += 3
    metrics = [
        ("Total Cable Pathways Assessed", ""),
        ("Pathways Compliant", ""),
        ("Pathways with Partial Compliance", ""),
        ("Pathways Non-Compliant", ""),
        ("Access-Controlled Wiring Closets", ""),
        ("Documentation Current", ""),
    ]

    for metric, value in metrics:
        ws.cell(row=row, column=1, value=metric).border = styles["border"]
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]
        row += 1

    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    ws[f"A{row+2}"] = "Gap List"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    row += 3
    gap_headers = ["Gap ID", "Description", "Priority", "Status", "Owner"]
    for col_idx, header in enumerate(gap_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_priority)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Closed,Accepted"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(row + 1, row + 21):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        dv_priority.add(ws.cell(row=r, column=3))
        dv_status.add(ws.cell(row=r, column=4))

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20

# =============================================================================
# EQUIPMENT MAINTENANCE SHEET
# =============================================================================

def create_equipment_maintenance_sheet(ws, styles):
    """Create Equipment Maintenance detail sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "A.7.13 Equipment Maintenance Compliance"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A3"] = "Score Summary"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    row = 4
    headers = ["Domain", "Weight", "Score", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    domains = [
        ("Schedule Compliance", "40%"),
        ("Maintenance Programme", "30%"),
        ("Personnel Verification", "15%"),
        ("Security Controls", "15%"),
        ("DOMAIN TOTAL", "100%"),
    ]

    row += 1
    for domain, weight in domains:
        ws.cell(row=row, column=1, value=domain)
        if domain == "DOMAIN TOTAL":
            ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True)
        ws.cell(row=row, column=1).border = styles["border"]

        ws.cell(row=row, column=2, value=weight).border = styles["border"]

        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=3).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=3).border = styles["border"]

        ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=4).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=4).border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Schedule Compliance Metrics"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    row += 3
    metrics = [
        ("Total Equipment in Programme", ""),
        ("Equipment Current", ""),
        ("Equipment Due Soon", ""),
        ("Equipment Overdue", ""),
        ("Critical Equipment Overdue", ""),
        ("Average Days Overdue", ""),
    ]

    for metric, value in metrics:
        ws.cell(row=row, column=1, value=metric).border = styles["border"]
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=2).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=2).border = styles["border"]
        row += 1

    ws[f"A{row+2}"] = "Overdue Equipment"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    row += 3
    overdue_headers = ["Equipment ID", "Description", "Criticality", "Days Overdue", "Escalated"]
    for col_idx, header in enumerate(overdue_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    for r in range(row + 1, row + 21):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12

# =============================================================================
# GAP REGISTER SHEET
# =============================================================================

def create_gap_register_sheet(ws, styles):
    """Create Gap Register sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "GAP REGISTER\nCentral tracking of all compliance gaps"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    columns = [
        ("Gap ID", 12),
        ("Source", 15),
        ("Description", 50),
        ("Priority", 12),
        ("Owner", 20),
        ("Identified Date", 15),
        ("Target Date", 15),
        ("Status", 15),
        ("Remediation Plan", 45),
        ("Evidence", 25),
        ("Closure Date", 15),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_source = DataValidation(
        type="list",
        formula1='"S1 - Cabling,S2 - Maintenance,S3 - Schedule"',
        allow_blank=True
    )
    ws.add_data_validation(dv_source)

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_priority)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Closed,Accepted"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"GAP-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_source.add(ws.cell(row=r, column=2))
        dv_priority.add(ws.cell(row=r, column=4))
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"

# =============================================================================
# TREND ANALYSIS SHEET
# =============================================================================

def create_trend_analysis_sheet(ws, styles):
    """Create Trend Analysis sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "TREND ANALYSIS\nHistorical compliance tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Monthly Compliance Data"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    columns = [
        ("Month", 15),
        ("Overall %", 12),
        ("Cabling %", 12),
        ("Maintenance %", 15),
        ("Open Gaps", 12),
        ("Notes", 40),
    ]

    row = 4
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate months
    months = [
        "Jan 2026", "Feb 2026", "Mar 2026", "Apr 2026", "May 2026", "Jun 2026",
        "Jul 2026", "Aug 2026", "Sep 2026", "Oct 2026", "Nov 2026", "Dec 2026",
    ]

    for r, month in enumerate(months, start=5):
        ws.cell(row=r, column=1, value=month).border = styles["border"]
        for c in range(2, 7):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

# =============================================================================
# AUDIT EVIDENCE SHEET
# =============================================================================

def create_audit_evidence_sheet(ws, styles):
    """Create Audit Evidence summary sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "AUDIT EVIDENCE SUMMARY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A3"] = "Assessment Evidence"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    columns = [
        ("Assessment", 35),
        ("File Name", 45),
        ("Location", 50),
        ("Last Updated", 15),
        ("Status", 15),
    ]

    row = 4
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    assessments = [
        ("S1 - Cabling Security", "ISMS-IMP-A.7.12-13.S1_Cabling_Security_*.xlsx"),
        ("S2 - Equipment Maintenance", "ISMS-IMP-A.7.12-13.S2_Equipment_Maintenance_*.xlsx"),
        ("S3 - Maintenance Schedule", "ISMS-IMP-A.7.12-13.S3_Maintenance_Schedule_*.xlsx"),
    ]

    dv_status = DataValidation(
        type="list",
        formula1='"Current,Outdated,Missing"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    row = 5
    for assessment, filename in assessments:
        ws.cell(row=row, column=1, value=assessment).border = styles["border"]
        ws.cell(row=row, column=2, value=filename).border = styles["border"]

        for c in range(3, 6):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        dv_status.add(ws.cell(row=row, column=5))
        row += 1

    ws[f"A{row+2}"] = "Supporting Evidence"
    ws[f"A{row+2}"].font = Font(name="Calibri", size=12, bold=True)

    row += 3
    ws.cell(row=row, column=1, value="See Evidence Registers in S1, S2, and S3 assessments for detailed evidence documentation.")
    ws.cell(row=row, column=1).font = Font(name="Calibri", italic=True)

# =============================================================================
# MAIN WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    styles = setup_styles()

    ws = wb.create_sheet("Executive Summary", 0)
    create_executive_summary_sheet(ws, styles)

    ws = wb.create_sheet("Cabling Security")
    create_cabling_security_sheet(ws, styles)

    ws = wb.create_sheet("Equipment Maintenance")
    create_equipment_maintenance_sheet(ws, styles)

    ws = wb.create_sheet("Gap Register")
    create_gap_register_sheet(ws, styles)

    ws = wb.create_sheet("Trend Analysis")
    create_trend_analysis_sheet(ws, styles)

    ws = wb.create_sheet("Audit Evidence")
    create_audit_evidence_sheet(ws, styles)

    return wb

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main entry point for workbook generation."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS Infrastructure Compliance Dashboard (A.7.12-13)")
        logger.info("=" * 70)

        wb = create_workbook()
        wb.save(OUTPUT_FILENAME)

        logger.info("%s SUCCESS: %s", CHECK, OUTPUT_FILENAME)
        logger.info("  %s 6 executive dashboard worksheets created", BULLET)
        logger.info("  %s Executive summary with overall compliance", BULLET)
        logger.info("  %s Cabling and Maintenance domain breakdowns", BULLET)
        logger.info("  %s Central gap register with 100 rows", BULLET)
        logger.info("  %s 12-month trend analysis", BULLET)
        logger.info("  %s Audit evidence summary", BULLET)
        logger.info("=" * 70)

        return 0

    except Exception as e:
        logger.error("%s ERROR: Failed to generate workbook", XMARK)
        logger.error("Error: %s", str(e))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
