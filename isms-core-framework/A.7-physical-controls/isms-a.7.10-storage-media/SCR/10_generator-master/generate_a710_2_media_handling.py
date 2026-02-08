#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.10.2 - Media Handling Procedures Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.10: Storage Media
Assessment Domain 2 of 4: Media Handling, Transportation & Access Controls

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates an Excel assessment workbook for evaluating media handling
procedures including access controls, transportation security, physical storage,
use procedures, and incident response.

**Generated Workbook Structure:**
1. Instructions & Legend
2. Media Access Controls
3. Transportation Security
4. Physical Storage Controls
5. Media Use Procedures
6. Incident Response
7. Summary Dashboard
8. Evidence Register
9. Approval Sign-Off

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

    python3 generate_a710_2_media_handling.py

================================================================================
"""

# =============================================================================
# Imports
# =============================================================================
import logging
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.10.2"
WORKBOOK_NAME = "Media Handling Procedures"
CONTROL_ID = "A.7.10"
CONTROL_NAME = "Storage Media"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
STYLES = {
    'title': {
        'font': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    },
    'header': {
        'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    },
    'subheader': {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D8E4F8', end_color='D8E4F8', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    },
    'input_cell': {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    }
}


def apply_style(cell, style_name):
    """Apply a named style to a cell."""
    style = STYLES[style_name]
    cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    cell.alignment = style['alignment']
    if 'border' in style:
        cell.border = style['border']


def create_workbook():
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions & Legend",
        "2. Media Access Controls",
        "3. Transportation Security",
        "4. Physical Storage Controls",
        "5. Media Use Procedures",
        "6. Incident Response",
        "Summary Dashboard",
        "Evidence Register",
        "Approval Sign-Off"
    ]

    for idx, sheet_name in enumerate(sheets):
        wb.create_sheet(sheet_name, idx)

    return wb


def get_base_columns():
    """Return base column structure."""
    return [
        ("A", "Control Area / Location", 25),
        ("B", "Classification Level", 18),
        ("C", "Control Owner", 18),
        ("D", "Procedure Reference", 20),
        ("E", "Last Review Date", 12),
        ("F", "Status", 18),
        ("G", "Implementation Date", 12),
        ("H", "Last Audit Date", 12),
        ("I", "Next Review Date", 12),
        ("J", "Gap Identified", 25),
        ("K", "Remediation Plan", 25),
        ("L", "Target Completion", 12),
        ("M", "Risk Level", 12),
        ("N", "Evidence Reference", 20),
        ("O", "Notes / Comments", 25),
        ("P", "Remediation Owner", 18),
        ("Q", "Budget Required", 15)
    ]


def create_base_validations(ws):
    """Create data validation objects for standard columns."""
    dv_classification = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Restricted,All Classifications"',
        allow_blank=True
    )
    ws.add_data_validation(dv_classification)
    dv_classification.add('B10:B100')

    dv_status = DataValidation(
        type="list",
        formula1='"Compliant,Partial,Non-Compliant,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add('F10:F100')

    dv_risk = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_risk)
    dv_risk.add('M10:M100')

    dv_budget = DataValidation(
        type="list",
        formula1='"Yes,No,Unknown"',
        allow_blank=True
    )
    ws.add_data_validation(dv_budget)
    dv_budget.add('Q10:Q100')


def create_assessment_sheet(ws, sheet_title, sheet_objective, extended_cols):
    """Create a standardized assessment sheet."""
    ws.merge_cells('A1:T1')
    cell = ws['A1']
    cell.value = sheet_title
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:T2')
    ws['A2'].value = "Policy Reference: ISMS-POL-A.7.10 - Storage Media"
    ws['A2'].font = Font(name='Calibri', size=9, italic=True)

    ws.merge_cells('A4:T6')
    cell = ws['A4']
    cell.value = f"ASSESSMENT OBJECTIVE:\n{sheet_objective}"
    cell.font = Font(name='Calibri', size=11, bold=True)
    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[4].height = 50

    ws.merge_cells('A7:T7')
    ws['A7'].value = "Complete the yellow-highlighted cells below. Use dropdowns where provided."
    ws['A7'].font = Font(name='Calibri', size=9, italic=True, color='0000FF')

    base_cols = get_base_columns()
    all_cols = base_cols + extended_cols

    for col_letter, header_text, width in all_cols:
        cell = ws[f'{col_letter}9']
        cell.value = header_text
        apply_style(cell, 'header')
        ws.column_dimensions[col_letter].width = width

    for row in range(10, 23):
        for col_idx in range(1, len(all_cols) + 1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f'{col_letter}{row}']
            apply_style(cell, 'input_cell')

    create_base_validations(ws)
    ws.freeze_panes = 'A10'


def create_instructions_sheet(ws):
    """Create Instructions & Legend sheet."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = f"{DOCUMENT_ID} - {WORKBOOK_NAME} Assessment"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    current_row = 3
    info_data = [
        ("Document ID:", DOCUMENT_ID),
        ("Assessment Date:", GENERATED_DATE),
        ("Assessor Name:", "[Enter Name]"),
        ("Organisation:", "[Enter Organisation]"),
        ("Control Reference:", CONTROL_REF)
    ]

    for label, value in info_data:
        ws[f'A{current_row}'].value = label
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'].value = value
        if "[" in value:
            ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    current_row += 1
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = "ASSESSMENT SCOPE"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    scope_items = [
        "- Media Access Controls: Who can access media, under what conditions",
        "- Transportation Security: Secure courier, personal transport, chain of custody",
        "- Physical Storage Controls: Secure cabinets, environmental conditions",
        "- Media Use Procedures: Data transfer, scanning, return processes",
        "- Incident Response: Lost/stolen media procedures, escalation"
    ]

    for item in scope_items:
        ws[f'A{current_row}'].value = item
        current_row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60


def create_access_controls_sheet(ws):
    """Create Media Access Controls sheet."""
    extended_cols = [
        ("R", "Access Control Type", 22),
        ("S", "Authorised Personnel", 25),
        ("T", "Review Frequency", 18)
    ]

    create_assessment_sheet(
        ws,
        "2. Media Access Controls",
        "Document access control requirements by media classification level including physical access, personnel authorisation, and review frequency.",
        extended_cols
    )

    dv_access = DataValidation(
        type="list",
        formula1='"Key Lock,Card Access,Biometric,Combination Lock,Multi-Factor,PIN Code,None"',
        allow_blank=True
    )
    ws.add_data_validation(dv_access)
    dv_access.add('R10:R100')

    dv_review = DataValidation(
        type="list",
        formula1='"Monthly,Quarterly,Semi-annual,Annual,Ad-hoc"',
        allow_blank=True
    )
    ws.add_data_validation(dv_review)
    dv_review.add('T10:T100')


def create_transportation_sheet(ws):
    """Create Transportation Security sheet."""
    extended_cols = [
        ("R", "Transport Method", 22),
        ("S", "Chain of Custody Required", 20),
        ("T", "Packaging Requirement", 25)
    ]

    create_assessment_sheet(
        ws,
        "3. Transportation Security",
        "Assess security controls for media in transit including courier requirements, chain of custody documentation, and packaging standards.",
        extended_cols
    )

    dv_transport = DataValidation(
        type="list",
        formula1='"Secure Courier,Standard Courier,Personal Transport,Internal Mail,Hand Delivery,Digital Transfer,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_transport)
    dv_transport.add('R10:R100')

    dv_custody = DataValidation(
        type="list",
        formula1='"Yes - Full Documentation,Yes - Basic Log,No,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_custody)
    dv_custody.add('S10:S100')

    dv_packaging = DataValidation(
        type="list",
        formula1='"Tamper-Evident Packaging,Sealed Container,Locked Case,Standard Packaging,N/A"',
        allow_blank=True
    )
    ws.add_data_validation(dv_packaging)
    dv_packaging.add('T10:T100')


def create_physical_storage_sheet(ws):
    """Create Physical Storage Controls sheet."""
    extended_cols = [
        ("R", "Storage Type", 22),
        ("S", "Environmental Monitoring", 22),
        ("T", "Fire Suppression", 20)
    ]

    create_assessment_sheet(
        ws,
        "4. Physical Storage Controls",
        "Assess physical storage security and environmental controls for media storage locations including temperature, humidity, and fire protection.",
        extended_cols
    )

    dv_storage = DataValidation(
        type="list",
        formula1='"Locked Cabinet,Fire-Rated Safe,Secure Vault,Server Room,Archive Room,Standard Shelf,Secure Disposal Bin"',
        allow_blank=True
    )
    ws.add_data_validation(dv_storage)
    dv_storage.add('R10:R100')

    dv_environmental = DataValidation(
        type="list",
        formula1='"Temperature Only,Humidity Only,Temperature + Humidity,Full Environmental,None"',
        allow_blank=True
    )
    ws.add_data_validation(dv_environmental)
    dv_environmental.add('S10:S100')

    dv_fire = DataValidation(
        type="list",
        formula1='"Gas Suppression,Wet Sprinkler,Dry Sprinkler,Fire Extinguisher,None"',
        allow_blank=True
    )
    ws.add_data_validation(dv_fire)
    dv_fire.add('T10:T100')


def create_use_procedures_sheet(ws):
    """Create Media Use Procedures sheet."""
    extended_cols = [
        ("R", "Applicable Media Type", 22),
        ("S", "Procedure Owner", 20),
        ("T", "Training Required", 18)
    ]

    create_assessment_sheet(
        ws,
        "5. Media Use Procedures",
        "Assess operational procedures for media use including data transfer, malware scanning, and return/disposal processes.",
        extended_cols
    )

    dv_media = DataValidation(
        type="list",
        formula1='"Removable Media,Fixed Storage,Cloud Storage,Paper Documents,All Media Types"',
        allow_blank=True
    )
    ws.add_data_validation(dv_media)
    dv_media.add('R10:R100')

    dv_training = DataValidation(
        type="list",
        formula1='"Yes - Initial Only,Yes - Annual Refresh,Yes - Quarterly,No,Role-Specific"',
        allow_blank=True
    )
    ws.add_data_validation(dv_training)
    dv_training.add('T10:T100')


def create_incident_response_sheet(ws):
    """Create Incident Response sheet."""
    extended_cols = [
        ("R", "Response Procedure", 25),
        ("S", "Escalation Path", 25),
        ("T", "Notification Required", 22)
    ]

    create_assessment_sheet(
        ws,
        "6. Incident Response",
        "Assess lost/stolen media procedures and incident response capabilities including escalation paths and notification requirements.",
        extended_cols
    )

    dv_notification = DataValidation(
        type="list",
        formula1='"Regulatory (GDPR/FINMA),Internal Only,Both Internal + Regulatory,None,To Be Determined"',
        allow_blank=True
    )
    ws.add_data_validation(dv_notification)
    dv_notification.add('T10:T100')


def create_summary_dashboard(ws):
    """Create Summary Dashboard sheet."""
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "Media Handling Procedures - Summary Dashboard"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    current_row = 3
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "CONTROL STATUS OVERVIEW"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    controls = [
        ("Access Controls Assessed", "[Count]"),
        ("Transportation Controls Assessed", "[Count]"),
        ("Storage Controls Assessed", "[Count]"),
        ("Use Procedures Assessed", "[Count]"),
        ("Incident Procedures Assessed", "[Count]")
    ]

    for label, value in controls:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    current_row += 1
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "COMPLIANCE METRICS"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    metrics = [
        ("Procedures with Documentation (%)", "[%]"),
        ("Procedures with Training (%)", "[%]"),
        ("Procedures Tested/Exercised (%)", "[%]"),
        ("Chain of Custody Compliance (%)", "[%]"),
        ("Approved Courier Usage (%)", "[%]")
    ]

    for label, value in metrics:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15


def create_evidence_register(ws):
    """Create Evidence Register sheet."""
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "Evidence Register - Supporting Documentation"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Evidence ID", 12),
        ("B", "Assessment Sheet", 20),
        ("C", "Related Control", 30),
        ("D", "Evidence Type", 20),
        ("E", "Evidence Description", 35),
        ("F", "File Location/Link", 40),
        ("G", "Date Collected", 12),
        ("H", "Retention Period", 15),
        ("I", "Next Review", 12),
        ("J", "Owner", 20),
        ("K", "Notes", 30)
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f'{col_letter}4']
        cell.value = header_text
        apply_style(cell, 'header')
        ws.column_dimensions[col_letter].width = width

    for row in range(5, 105):
        for col_letter, _, _ in headers:
            cell = ws[f'{col_letter}{row}']
            apply_style(cell, 'input_cell')

    ws.freeze_panes = 'A5'


def create_approval_signoff(ws):
    """Create Approval Sign-Off sheet."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "Three-Level Approval Workflow"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    current_row = 3
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = "DOCUMENT CONTROL"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    fields = [
        ("Assessment Period:", "[e.g., Q1 2026]"),
        ("Workbook Version:", "1.0"),
        ("Controls Assessed:", "[Count]"),
        ("Overall Compliance %:", "[%]"),
        ("Assessment Completed By:", "[Name, Date]")
    ]

    for label, value in fields:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    current_row += 2

    approvals = [
        ("LEVEL 1 APPROVAL - Technical/Operational", "Physical Security Manager / IT Operations Manager"),
        ("LEVEL 2 APPROVAL - Management", "CISO / COO"),
        ("LEVEL 3 APPROVAL - Executive", "Executive Management / CRO")
    ]

    approval_fields = ["Approver Name:", "Title:", "Date:", "Status:", "Comments:", "Signature:"]

    for level_title, role in approvals:
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'].value = level_title
        apply_style(ws[f'A{current_row}'], 'subheader')
        current_row += 1

        ws[f'A{current_row}'].value = f"Role: {role}"
        ws[f'A{current_row}'].font = Font(italic=True)
        current_row += 1

        for field in approval_fields:
            ws[f'A{current_row}'].value = field
            ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            if field == "Status:":
                dv = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected"', allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(f'B{current_row}')
            current_row += 1

        current_row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50


def main():
    """Main function to generate the workbook."""
    logger.info(f"Starting generation of {DOCUMENT_ID} - {WORKBOOK_NAME}")

    try:
        wb = create_workbook()

        logger.info("Creating Instructions & Legend sheet...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("Creating Media Access Controls sheet...")
        create_access_controls_sheet(wb["2. Media Access Controls"])

        logger.info("Creating Transportation Security sheet...")
        create_transportation_sheet(wb["3. Transportation Security"])

        logger.info("Creating Physical Storage Controls sheet...")
        create_physical_storage_sheet(wb["4. Physical Storage Controls"])

        logger.info("Creating Media Use Procedures sheet...")
        create_use_procedures_sheet(wb["5. Media Use Procedures"])

        logger.info("Creating Incident Response sheet...")
        create_incident_response_sheet(wb["6. Incident Response"])

        logger.info("Creating Summary Dashboard...")
        create_summary_dashboard(wb["Summary Dashboard"])

        logger.info("Creating Evidence Register...")
        create_evidence_register(wb["Evidence Register"])

        logger.info("Creating Approval Sign-Off...")
        create_approval_signoff(wb["Approval Sign-Off"])

        logger.info(f"Saving workbook as {OUTPUT_FILENAME}...")
        wb.save(OUTPUT_FILENAME)

        logger.info("=" * 60)
        logger.info(f"SUCCESS: Workbook generated successfully!")
        logger.info(f"Output file: {OUTPUT_FILENAME}")
        logger.info("=" * 60)

        return True

    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation following A.8.10 pattern
# =============================================================================
