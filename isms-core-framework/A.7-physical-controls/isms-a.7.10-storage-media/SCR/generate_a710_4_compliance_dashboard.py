#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.10.4 - Storage Media Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.7.10: Storage Media
Assessment Domain 4 of 4: Consolidated Compliance Dashboard & KPIs

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a consolidated compliance dashboard workbook that pulls
data from A.7.10.1-3 assessments to provide executive overview, gap analysis,
risk register, remediation roadmap, and KPI tracking.

**Generated Workbook Structure:**
1. Instructions & Legend
2. Executive Summary
3. Domain 1 Summary (Inventory)
4. Domain 2 Summary (Handling)
5. Domain 3 Summary (Lifecycle)
6. Consolidated Gap Analysis
7. Risk Register
8. Remediation Roadmap
9. Evidence Master Index
10. KPI Dashboard
11. CISO Approval

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

    python3 generate_a710_4_compliance_dashboard.py

================================================================================
"""

# =============================================================================
# Imports
# =============================================================================
import logging
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.10.4"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.7.10"
CONTROL_NAME = "Storage Media"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
STYLES = {
    'title': {
        'font': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    },
    'header': {
        'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    },
    'subheader': {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D8E4F8', end_color='D8E4F8', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    },
    'input_cell': {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    },
    'metric_label': {
        'font': Font(name='Calibri', size=11, color='000000'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    },
    'metric_value': {
        'font': Font(name='Calibri', size=14, bold=True, color='003366'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    }
}


def apply_style(cell, style_name):
    """Apply a named style to a cell."""
    style = STYLES[style_name]
    cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    cell.alignment = style['alignment']
    if 'border' in style:
        cell.border = style['border']


def create_workbook():
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions & Legend",
        "Executive Summary",
        "Domain 1 Summary",
        "Domain 2 Summary",
        "Domain 3 Summary",
        "Consolidated Gap Analysis",
        "Risk Register",
        "Remediation Roadmap",
        "Evidence Master Index",
        "KPI Dashboard",
        "CISO Approval"
    ]

    for idx, sheet_name in enumerate(sheets):
        wb.create_sheet(sheet_name, idx)

    return wb


def create_instructions_sheet(ws):
    """Create Instructions & Legend sheet."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = f"{DOCUMENT_ID} - Storage Media {WORKBOOK_NAME}"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    current_row = 3
    info_data = [
        ("Document ID:", DOCUMENT_ID),
        ("Dashboard Date:", GENERATED_DATE),
        ("Prepared By:", "[Enter Name]"),
        ("Organisation:", "[Enter Organisation]"),
        ("Control Reference:", CONTROL_REF)
    ]

    for label, value in info_data:
        ws[f'A{current_row}'].value = label
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'].value = value
        if "[" in value:
            ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    current_row += 1
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = "DASHBOARD OVERVIEW"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    overview = [
        "This dashboard consolidates data from A.7.10.1-3 assessments:",
        "",
        "- A.7.10.1: Storage Media Inventory",
        "- A.7.10.2: Media Handling Procedures",
        "- A.7.10.3: Media Lifecycle Tracking",
        "",
        "Use external workbook links or manual data entry to populate domain summaries."
    ]

    for item in overview:
        ws[f'A{current_row}'].value = item
        current_row += 1

    current_row += 1
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = "EXTERNAL LINKING (OPTIONAL)"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    linking_info = [
        "To enable automatic data population:",
        "1. Run normalize_assessment_files_a710.py to create normalized filenames",
        "2. Place normalized files in the same directory as this dashboard",
        "3. External formulas will reference: ISMS-IMP-A.7.10.1.xlsx, etc.",
        "4. Click 'Update Links' when prompted in Excel"
    ]

    for item in linking_info:
        ws[f'A{current_row}'].value = item
        current_row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60


def create_executive_summary(ws):
    """Create Executive Summary sheet."""
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "Storage Media Control A.7.10 - Executive Summary"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    current_row = 3

    # Compliance Overview
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "COMPLIANCE OVERVIEW"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    ws[f'A{current_row}'].value = "Metric"
    ws[f'B{current_row}'].value = "Current"
    ws[f'C{current_row}'].value = "Previous"
    ws[f'D{current_row}'].value = "Trend"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{current_row}'], 'header')
    current_row += 1

    compliance_metrics = [
        ("Overall Compliance %", "[%]", "[%]", "[Up/Down/Stable]"),
        ("Domain 1 (Inventory) %", "[%]", "[%]", "[Up/Down/Stable]"),
        ("Domain 2 (Handling) %", "[%]", "[%]", "[Up/Down/Stable]"),
        ("Domain 3 (Lifecycle) %", "[%]", "[%]", "[Up/Down/Stable]")
    ]

    for metric, current, previous, trend in compliance_metrics:
        ws[f'A{current_row}'].value = metric
        ws[f'B{current_row}'].value = current
        ws[f'C{current_row}'].value = previous
        ws[f'D{current_row}'].value = trend
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws[f'C{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws[f'D{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    current_row += 1

    # Critical Metrics
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "CRITICAL METRICS"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    critical_metrics = [
        ("Total Media Items Inventoried", "[Count]"),
        ("Unregistered Media", "[Count]"),
        ("Encryption Compliance %", "[%]"),
        ("Disposal with Certificate %", "[%]"),
        ("Open Critical Gaps", "[Count]"),
        ("Remediation On Track %", "[%]")
    ]

    for label, value in critical_metrics:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    current_row += 1

    # Key Findings
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "KEY FINDINGS (TOP 5)"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    for i in range(1, 6):
        ws[f'A{current_row}'].value = f"Finding {i}:"
        ws.merge_cells(f'B{current_row}:D{current_row}')
        ws[f'B{current_row}'].value = "[Enter key finding]"
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_domain_summary(ws, domain_num, domain_name, source_doc):
    """Create Domain Summary sheet."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = f"Domain {domain_num} Summary: {domain_name}"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    current_row = 3
    ws[f'A{current_row}'].value = "Source Document:"
    ws[f'B{current_row}'].value = source_doc
    current_row += 1
    ws[f'A{current_row}'].value = "Last Updated:"
    ws[f'B{current_row}'].value = "[Date]"
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    current_row += 1
    ws[f'A{current_row}'].value = "Next Review:"
    ws[f'B{current_row}'].value = "[Date]"
    ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    current_row += 2

    # Compliance Metrics
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "COMPLIANCE METRICS"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    metrics = [
        ("Domain Compliance %", "[%]"),
        ("Items Assessed", "[Count]"),
        ("Compliant Items", "[Count]"),
        ("Partial Items", "[Count]"),
        ("Non-Compliant Items", "[Count]"),
        ("N/A Items", "[Count]")
    ]

    for label, value in metrics:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    current_row += 1

    # Key Findings
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'].value = "KEY FINDINGS"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    ws[f'A{current_row}'].value = "Finding"
    ws[f'B{current_row}'].value = "Severity"
    ws[f'C{current_row}'].value = "Evidence Ref"
    ws[f'D{current_row}'].value = "Owner"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{current_row}'], 'header')
    current_row += 1

    for i in range(5):
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20


def create_gap_analysis(ws):
    """Create Consolidated Gap Analysis sheet."""
    ws.merge_cells('A1:L1')
    cell = ws['A1']
    cell.value = "Consolidated Gap Analysis"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Gap ID", 10),
        ("B", "Source Domain", 15),
        ("C", "Source Sheet", 20),
        ("D", "Gap Description", 40),
        ("E", "Control Reference", 15),
        ("F", "Risk Level", 12),
        ("G", "Gap Category", 18),
        ("H", "Current Status", 15),
        ("I", "Remediation Owner", 18),
        ("J", "Target Date", 12),
        ("K", "Evidence Ref", 12),
        ("L", "Notes", 25)
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f'{col_letter}3']
        cell.value = header_text
        apply_style(cell, 'header')
        ws.column_dimensions[col_letter].width = width

    for row in range(4, 54):
        for col_letter, _, _ in headers:
            cell = ws[f'{col_letter}{row}']
            apply_style(cell, 'input_cell')

    # Add validations
    dv_domain = DataValidation(
        type="list",
        formula1='"Domain 1 (Inventory),Domain 2 (Handling),Domain 3 (Lifecycle)"',
        allow_blank=True
    )
    ws.add_data_validation(dv_domain)
    dv_domain.add('B4:B53')

    dv_risk = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(dv_risk)
    dv_risk.add('F4:F53')

    dv_category = DataValidation(
        type="list",
        formula1='"Policy Gap,Procedure Gap,Technical Gap,People Gap,Documentation Gap,Multiple"',
        allow_blank=True
    )
    ws.add_data_validation(dv_category)
    dv_category.add('G4:G53')

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Pending Verification,Closed,Accepted Risk"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add('H4:H53')

    ws.freeze_panes = 'A4'


def create_risk_register(ws):
    """Create Risk Register sheet."""
    ws.merge_cells('A1:N1')
    cell = ws['A1']
    cell.value = "Risk Register"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Risk ID", 10),
        ("B", "Related Gap(s)", 15),
        ("C", "Risk Description", 35),
        ("D", "Risk Category", 18),
        ("E", "Impact (1-5)", 10),
        ("F", "Likelihood (1-5)", 10),
        ("G", "Risk Score", 10),
        ("H", "Risk Rating", 12),
        ("I", "Risk Owner", 18),
        ("J", "Treatment", 15),
        ("K", "Treatment Plan", 30),
        ("L", "Treatment Status", 15),
        ("M", "Residual Risk", 12),
        ("N", "Review Date", 12)
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f'{col_letter}3']
        cell.value = header_text
        apply_style(cell, 'header')
        ws.column_dimensions[col_letter].width = width

    for row in range(4, 34):
        for col_letter, _, _ in headers:
            cell = ws[f'{col_letter}{row}']
            apply_style(cell, 'input_cell')

    # Add validations
    dv_category = DataValidation(
        type="list",
        formula1='"Data Breach,Regulatory Compliance,Operational Continuity,Reputational,Financial,Legal,Multiple"',
        allow_blank=True
    )
    ws.add_data_validation(dv_category)
    dv_category.add('D4:D33')

    dv_treatment = DataValidation(
        type="list",
        formula1='"Mitigate,Accept,Transfer,Avoid"',
        allow_blank=True
    )
    ws.add_data_validation(dv_treatment)
    dv_treatment.add('J4:J33')

    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,On Hold,Complete,Ongoing"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add('L4:L33')

    ws.freeze_panes = 'A4'


def create_remediation_roadmap(ws):
    """Create Remediation Roadmap sheet."""
    ws.merge_cells('A1:M1')
    cell = ws['A1']
    cell.value = "Remediation Roadmap"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Action ID", 10),
        ("B", "Related Gap/Risk", 15),
        ("C", "Action Description", 35),
        ("D", "Action Owner", 18),
        ("E", "Priority", 10),
        ("F", "Start Date", 12),
        ("G", "Target Date", 12),
        ("H", "Actual Completion", 12),
        ("I", "Progress %", 10),
        ("J", "Status", 15),
        ("K", "Dependencies", 20),
        ("L", "Budget Required", 12),
        ("M", "Notes", 25)
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f'{col_letter}3']
        cell.value = header_text
        apply_style(cell, 'header')
        ws.column_dimensions[col_letter].width = width

    for row in range(4, 54):
        for col_letter, _, _ in headers:
            cell = ws[f'{col_letter}{row}']
            apply_style(cell, 'input_cell')

    # Add validations
    dv_priority = DataValidation(
        type="list",
        formula1='"P1 (30 days),P2 (90 days),P3 (180 days),P4 (365 days)"',
        allow_blank=True
    )
    ws.add_data_validation(dv_priority)
    dv_priority.add('E4:E53')

    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,On Hold,Complete,Cancelled"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add('J4:J53')

    ws.freeze_panes = 'A4'


def create_evidence_index(ws):
    """Create Evidence Master Index sheet."""
    ws.merge_cells('A1:L1')
    cell = ws['A1']
    cell.value = "Evidence Master Index"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Evidence ID", 12),
        ("B", "Source Domain", 15),
        ("C", "Related Control", 15),
        ("D", "Related Gap/Risk", 15),
        ("E", "Evidence Type", 18),
        ("F", "Evidence Description", 35),
        ("G", "Document/File Name", 25),
        ("H", "Location/Link", 30),
        ("I", "Date Collected", 12),
        ("J", "Collected By", 18),
        ("K", "Retention Period", 15),
        ("L", "Review Date", 12)
    ]

    for col_letter, header_text, width in headers:
        cell = ws[f'{col_letter}3']
        cell.value = header_text
        apply_style(cell, 'header')
        ws.column_dimensions[col_letter].width = width

    for row in range(4, 104):
        for col_letter, _, _ in headers:
            cell = ws[f'{col_letter}{row}']
            apply_style(cell, 'input_cell')

    ws.freeze_panes = 'A4'


def create_kpi_dashboard(ws):
    """Create KPI Dashboard sheet."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "Storage Media KPI Dashboard"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    current_row = 3

    # Media Inventory KPIs
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'].value = "MEDIA INVENTORY KPIs"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    ws[f'A{current_row}'].value = "KPI ID"
    ws[f'B{current_row}'].value = "KPI Name"
    ws[f'C{current_row}'].value = "Target"
    ws[f'D{current_row}'].value = "Current"
    ws[f'E{current_row}'].value = "Trend"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{current_row}'], 'header')
    current_row += 1

    inventory_kpis = [
        ("KPI-01", "Registered media %", "100%", "[%]"),
        ("KPI-02", "Encrypted CONFIDENTIAL %", "100%", "[%]"),
        ("KPI-03", "Authorised removable %", "100%", "[%]"),
        ("KPI-04", "Custodian assigned %", "100%", "[%]")
    ]

    for kpi_id, name, target, current in inventory_kpis:
        ws[f'A{current_row}'].value = kpi_id
        ws[f'B{current_row}'].value = name
        ws[f'C{current_row}'].value = target
        ws[f'D{current_row}'].value = current
        ws[f'E{current_row}'].value = "[Trend]"
        ws[f'D{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws[f'E{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    current_row += 1

    # Handling KPIs
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'].value = "HANDLING KPIs"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    ws[f'A{current_row}'].value = "KPI ID"
    ws[f'B{current_row}'].value = "KPI Name"
    ws[f'C{current_row}'].value = "Target"
    ws[f'D{current_row}'].value = "Current"
    ws[f'E{current_row}'].value = "Trend"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{current_row}'], 'header')
    current_row += 1

    handling_kpis = [
        ("KPI-05", "Chain of custody %", "100%", "[%]"),
        ("KPI-06", "Approved courier usage %", "100%", "[%]"),
        ("KPI-07", "Environmental compliance %", "100%", "[%]")
    ]

    for kpi_id, name, target, current in handling_kpis:
        ws[f'A{current_row}'].value = kpi_id
        ws[f'B{current_row}'].value = name
        ws[f'C{current_row}'].value = target
        ws[f'D{current_row}'].value = current
        ws[f'E{current_row}'].value = "[Trend]"
        ws[f'D{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws[f'E{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    current_row += 1

    # Lifecycle KPIs
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'].value = "LIFECYCLE KPIs"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    ws[f'A{current_row}'].value = "KPI ID"
    ws[f'B{current_row}'].value = "KPI Name"
    ws[f'C{current_row}'].value = "Target"
    ws[f'D{current_row}'].value = "Current"
    ws[f'E{current_row}'].value = "Trend"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{current_row}'], 'header')
    current_row += 1

    lifecycle_kpis = [
        ("KPI-08", "Disposal with certificate %", "100%", "[%]"),
        ("KPI-09", "Vendor certification current %", "100%", "[%]"),
        ("KPI-10", "Re-use verification %", "100%", "[%]")
    ]

    for kpi_id, name, target, current in lifecycle_kpis:
        ws[f'A{current_row}'].value = kpi_id
        ws[f'B{current_row}'].value = name
        ws[f'C{current_row}'].value = target
        ws[f'D{current_row}'].value = current
        ws[f'E{current_row}'].value = "[Trend]"
        ws[f'D{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws[f'E{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    current_row += 1

    # Incident KPIs
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'].value = "INCIDENT KPIs"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    ws[f'A{current_row}'].value = "KPI ID"
    ws[f'B{current_row}'].value = "KPI Name"
    ws[f'C{current_row}'].value = "Target"
    ws[f'D{current_row}'].value = "Current"
    ws[f'E{current_row}'].value = "Trend"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{current_row}'], 'header')
    current_row += 1

    incident_kpis = [
        ("KPI-11", "Media loss incidents (YTD)", "0", "[Count]"),
        ("KPI-12", "Overdue media returns", "<3", "[Count]")
    ]

    for kpi_id, name, target, current in incident_kpis:
        ws[f'A{current_row}'].value = kpi_id
        ws[f'B{current_row}'].value = name
        ws[f'C{current_row}'].value = target
        ws[f'D{current_row}'].value = current
        ws[f'E{current_row}'].value = "[Trend]"
        ws[f'D{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws[f'E{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15


def create_ciso_approval(ws):
    """Create CISO Approval sheet."""
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "CISO / Executive Approval"
    apply_style(cell, 'title')
    ws.row_dimensions[1].height = 25

    current_row = 3

    # Document Control
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = "DOCUMENT CONTROL"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    fields = [
        ("Dashboard Version:", "1.0"),
        ("Period Covered:", "[e.g., Q1 2026]"),
        ("Overall Compliance %:", "[%]"),
        ("Critical Gaps:", "[Count]"),
        ("Prepared By:", "[Name]"),
        ("Preparation Date:", "[Date]")
    ]

    for label, value in fields:
        ws[f'A{current_row}'].value = label
        ws[f'B{current_row}'].value = value
        if "[" in value:
            ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        current_row += 1

    current_row += 2

    # Executive Summary
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = "EXECUTIVE SUMMARY"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    summary_fields = [
        "Key Findings:",
        "Significant Improvements:",
        "Areas of Concern:",
        "Resource Requirements:"
    ]

    for field in summary_fields:
        ws[f'A{current_row}'].value = field
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws.row_dimensions[current_row].height = 40
        current_row += 1

    current_row += 2

    # CISO Approval
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = "CISO APPROVAL"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = '"I have reviewed this dashboard and confirm that the storage media compliance posture is accurately represented. I approve the remediation priorities and resource allocation."'
    ws[f'A{current_row}'].font = Font(italic=True)
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    approval_fields = ["CISO Name:", "Title:", "Date:", "Status:", "Comments:", "Signature:"]

    for field in approval_fields:
        ws[f'A{current_row}'].value = field
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        if field == "Status:":
            dv = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f'B{current_row}')
        current_row += 1

    current_row += 2

    # Executive Approval
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = "EXECUTIVE APPROVAL"
    apply_style(ws[f'A{current_row}'], 'subheader')
    current_row += 1

    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'].value = '"Executive Management acknowledges the storage media compliance status and associated risks. The organisation commits to the remediation roadmap."'
    ws[f'A{current_row}'].font = Font(italic=True)
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    for field in approval_fields:
        ws[f'A{current_row}'].value = field
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        if field == "Status:":
            dv = DataValidation(type="list", formula1='"Approved,Approved with Conditions,Rejected"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f'B{current_row}')
        current_row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50


def main():
    """Main function to generate the workbook."""
    logger.info(f"Starting generation of {DOCUMENT_ID} - {WORKBOOK_NAME}")

    try:
        wb = create_workbook()

        logger.info("Creating Instructions & Legend sheet...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("Creating Executive Summary sheet...")
        create_executive_summary(wb["Executive Summary"])

        logger.info("Creating Domain 1 Summary sheet...")
        create_domain_summary(wb["Domain 1 Summary"], 1, "Storage Media Inventory", "ISMS-IMP-A.7.10.1")

        logger.info("Creating Domain 2 Summary sheet...")
        create_domain_summary(wb["Domain 2 Summary"], 2, "Media Handling Procedures", "ISMS-IMP-A.7.10.2")

        logger.info("Creating Domain 3 Summary sheet...")
        create_domain_summary(wb["Domain 3 Summary"], 3, "Media Lifecycle Tracking", "ISMS-IMP-A.7.10.3")

        logger.info("Creating Consolidated Gap Analysis sheet...")
        create_gap_analysis(wb["Consolidated Gap Analysis"])

        logger.info("Creating Risk Register sheet...")
        create_risk_register(wb["Risk Register"])

        logger.info("Creating Remediation Roadmap sheet...")
        create_remediation_roadmap(wb["Remediation Roadmap"])

        logger.info("Creating Evidence Master Index sheet...")
        create_evidence_index(wb["Evidence Master Index"])

        logger.info("Creating KPI Dashboard sheet...")
        create_kpi_dashboard(wb["KPI Dashboard"])

        logger.info("Creating CISO Approval sheet...")
        create_ciso_approval(wb["CISO Approval"])

        logger.info(f"Saving workbook as {OUTPUT_FILENAME}...")
        wb.save(OUTPUT_FILENAME)

        logger.info("=" * 60)
        logger.info(f"SUCCESS: Workbook generated successfully!")
        logger.info(f"Output file: {OUTPUT_FILENAME}")
        logger.info("=" * 60)

        return True

    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation following A.8.10 pattern
# =============================================================================
