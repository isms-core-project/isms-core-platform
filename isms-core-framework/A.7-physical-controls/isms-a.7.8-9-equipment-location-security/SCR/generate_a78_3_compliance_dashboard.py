#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.7.8-9.S3 - Equipment Protection Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.7.8 & A.7.9: Equipment Siting and Protection
Assessment Domain 3 of 3: Consolidated Compliance Dashboard

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific compliance dashboard requirements and reporting
needs.

Key customization areas:
1. Score weightings (adjust based on organizational priorities)
2. Gap priority definitions (match your risk framework)
3. Audit readiness checklist (adapt to your documentation requirements)
4. External workbook links (configure for your file locations)

Reference Pattern: Based on ISMS-A.7.4 Physical Infrastructure Framework

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Usage:
    python3 generate_a78_3_compliance_dashboard.py

Output:
    ISMS-IMP-A.7.8-9.S3_Compliance_Dashboard_YYYYMMDD.xlsx

--------------------------------------------------------------------------------
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.7.8-9.S3"
WORKBOOK_NAME = "Equipment Protection Compliance Dashboard"
CONTROL_ID = "A.7.8-9"
CONTROL_NAME = "Equipment Siting and Protection"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls A.7.8 & A.7.9: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID.replace('.', '_')}_Compliance_Dashboard_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# UNICODE SYMBOLS
# =============================================================================
CHECK = '\u2705'
XMARK = '\u274C'
WARNING = '\u26A0'
BULLET = '\u2022'
ARROW_UP = '\u2191'
ARROW_DOWN = '\u2193'

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "score_high": {
            "font": Font(name="Calibri", size=24, bold=True, color="006600"),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "score_medium": {
            "font": Font(name="Calibri", size=24, bold=True, color="9C5700"),
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "score_low": {
            "font": Font(name="Calibri", size=24, bold=True, color="9C0006"),
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles

# =============================================================================
# EXECUTIVE SUMMARY SHEET
# =============================================================================

def create_executive_summary(ws, styles):
    """Create Executive Summary dashboard sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = (
        "ISMS Equipment Protection Compliance Dashboard\n"
        "ISO/IEC 27001:2022 Controls A.7.8 & A.7.9"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 50

    # Overall Score Section
    ws.merge_cells("A3:B3")
    ws["A3"] = "OVERALL COMPLIANCE SCORE"
    ws["A3"].font = styles["section_header"]["font"]
    ws["A3"].fill = styles["section_header"]["fill"]
    ws["A3"].alignment = styles["section_header"]["alignment"]

    ws.merge_cells("A4:B6")
    ws["A4"] = "='Data Input'!C18"
    ws["A4"].font = Font(name="Calibri", size=36, bold=True)
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")

    ws["C4"] = "Status:"
    ws["C4"].font = Font(name="Calibri", bold=True)
    ws["D4"] = '=IF(VALUE(SUBSTITUTE(\'Data Input\'!C18,"%",""))>=90,"Compliant",IF(VALUE(SUBSTITUTE(\'Data Input\'!C18,"%",""))>=75,"Partial","Non-Compliant"))'

    ws["C5"] = "Last Updated:"
    ws["C5"].font = Font(name="Calibri", bold=True)
    ws["D5"].fill = styles["input_cell"]["fill"]
    ws["D5"].border = styles["border"]

    ws["C6"] = "Assessment Period:"
    ws["C6"].font = Font(name="Calibri", bold=True)
    ws["D6"].fill = styles["input_cell"]["fill"]
    ws["D6"].border = styles["border"]

    # Control Area Scores
    ws.merge_cells("A8:H8")
    ws["A8"] = "CONTROL AREA SCORES"
    ws["A8"].font = styles["section_header"]["font"]
    ws["A8"].fill = styles["section_header"]["fill"]
    ws["A8"].alignment = styles["section_header"]["alignment"]

    headers = ["Control", "Description", "Score", "Items", "Compliant", "Partial", "Non-Compliant", "Status"]
    row = 9
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    controls = [
        ("A.7.8", "Equipment Siting", "='Data Input'!C10", "='Data Input'!D10", "='Data Input'!E10", "='Data Input'!F10", "='Data Input'!G10"),
        ("A.7.9", "Off-Premises Security", "='Data Input'!C16", "='Data Input'!D16", "='Data Input'!E16", "='Data Input'!F16", "='Data Input'!G16"),
    ]

    row = 10
    for control, desc, score, items, comp, partial, noncomp in controls:
        ws.cell(row=row, column=1, value=control).font = Font(name="Calibri", bold=True)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=score)
        ws.cell(row=row, column=4, value=items)
        ws.cell(row=row, column=5, value=comp)
        ws.cell(row=row, column=6, value=partial)
        ws.cell(row=row, column=7, value=noncomp)
        ws.cell(row=row, column=8, value=f'=IF(VALUE(SUBSTITUTE(C{row},"%",""))>=90,"{CHECK}",IF(VALUE(SUBSTITUTE(C{row},"%",""))>=75,"{WARNING}","{XMARK}"))')
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = styles["border"]
        row += 1

    # Key Metrics
    ws.merge_cells("A13:H13")
    ws["A13"] = "KEY METRICS"
    ws["A13"].font = styles["section_header"]["font"]
    ws["A13"].fill = styles["section_header"]["fill"]
    ws["A13"].alignment = styles["section_header"]["alignment"]

    metrics = [
        ("Total Open Gaps:", "=COUNTIF('Gap Register'!J:J,\"Open\")+COUNTIF('Gap Register'!J:J,\"In Progress\")"),
        ("Critical Gaps:", "=COUNTIF('Gap Register'!F:F,\"Critical\")"),
        ("High Priority Gaps:", "=COUNTIF('Gap Register'!F:F,\"High\")"),
        ("Incidents This Quarter:", "=COUNTA('Incident Tracker'!A:A)-1"),
        ("Audit Readiness:", "='Audit Readiness'!C15"),
    ]

    row = 14
    for metric, formula in metrics:
        ws.cell(row=row, column=1, value=metric).font = Font(name="Calibri", bold=True)
        ws.cell(row=row, column=2, value=formula)
        row += 1

    # Critical Alerts
    ws.merge_cells("A20:H20")
    ws["A20"] = "CRITICAL ALERTS"
    ws["A20"].font = styles["section_header"]["font"]
    ws["A20"].fill = styles["section_header"]["fill"]
    ws["A20"].alignment = styles["section_header"]["alignment"]

    ws["A21"] = f"{XMARK} Review 'Gap Register' sheet for open critical and high priority gaps"
    ws["A22"] = f"{WARNING} Check 'Incident Tracker' for recent equipment security incidents"
    ws["A23"] = f"{CHECK} Verify 'Audit Readiness' checklist before scheduled audits"

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 10

# =============================================================================
# CONTROL AREA DETAILS SHEET
# =============================================================================

def create_control_details(ws, styles):
    """Create Control Area Details sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "CONTROL AREA DETAILS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # A.7.8 Equipment Siting Section
    ws.merge_cells("A3:G3")
    ws["A3"] = "A.7.8 - EQUIPMENT SITING AND PROTECTION"
    ws["A3"].font = styles["section_header"]["font"]
    ws["A3"].fill = styles["section_header"]["fill"]

    headers = ["Sub-domain", "Score", "Items", "Compliant", "Partial", "Non-Compliant", "N/A"]
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    a78_domains = [
        "Equipment Locations",
        "Environmental Assessment",
        "Physical Security",
        "Power Infrastructure",
        "Workstation Security",
    ]

    row = 5
    for i, domain in enumerate(a78_domains):
        ws.cell(row=row, column=1, value=domain)
        ws.cell(row=row, column=2, value=f"='Data Input'!C{5+i}")
        ws.cell(row=row, column=3, value=f"='Data Input'!D{5+i}")
        ws.cell(row=row, column=4, value=f"='Data Input'!E{5+i}")
        ws.cell(row=row, column=5, value=f"='Data Input'!F{5+i}")
        ws.cell(row=row, column=6, value=f"='Data Input'!G{5+i}")
        ws.cell(row=row, column=7, value=f"='Data Input'!H{5+i}")
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = styles["border"]
        row += 1

    # A.7.8 Total
    ws.cell(row=row, column=1, value="A.7.8 TOTAL").font = Font(name="Calibri", bold=True)
    ws.cell(row=row, column=2, value="='Data Input'!C10").font = Font(name="Calibri", bold=True)
    for c in range(1, 8):
        ws.cell(row=row, column=c).border = styles["border"]

    # A.7.9 Off-Premises Section
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "A.7.9 - SECURITY OF ASSETS OFF-PREMISES"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]

    row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    a79_domains = [
        "Equipment Inventory",
        "Authorisation & Tracking",
        "Protection Measures",
        "Remote Working",
        "Permanent Off-Site",
    ]

    row += 1
    for i, domain in enumerate(a79_domains):
        ws.cell(row=row, column=1, value=domain)
        ws.cell(row=row, column=2, value=f"='Data Input'!C{11+i}")
        ws.cell(row=row, column=3, value=f"='Data Input'!D{11+i}")
        ws.cell(row=row, column=4, value=f"='Data Input'!E{11+i}")
        ws.cell(row=row, column=5, value=f"='Data Input'!F{11+i}")
        ws.cell(row=row, column=6, value=f"='Data Input'!G{11+i}")
        ws.cell(row=row, column=7, value="0")
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = styles["border"]
        row += 1

    # A.7.9 Total
    ws.cell(row=row, column=1, value="A.7.9 TOTAL").font = Font(name="Calibri", bold=True)
    ws.cell(row=row, column=2, value="='Data Input'!C16").font = Font(name="Calibri", bold=True)
    for c in range(1, 8):
        ws.cell(row=row, column=c).border = styles["border"]

    # Column widths
    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 20

# =============================================================================
# GAP REGISTER SHEET
# =============================================================================

def create_gap_register(ws, styles):
    """Create Gap Register sheet."""
    ws.merge_cells("A1:N1")
    ws["A1"] = "GAP REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:N2")
    ws["A2"] = "Track all compliance gaps requiring remediation."
    ws["A2"].font = Font(name="Calibri", italic=True)

    columns = {
        "Gap ID": 12,
        "Control": 10,
        "Sub-domain": 25,
        "Description": 50,
        "Risk Impact": 40,
        "Priority": 12,
        "Owner": 20,
        "Identified Date": 14,
        "Target Date": 14,
        "Status": 15,
        "Days Open": 12,
        "Overdue": 10,
        "Evidence Ref": 15,
        "Notes": 40,
    }

    row = 4
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_control = DataValidation(
        type="list",
        formula1='"A.7.8,A.7.9"',
        allow_blank=False
    )
    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Completed,Deferred"',
        allow_blank=False
    )

    ws.add_data_validation(dv_control)
    ws.add_data_validation(dv_priority)
    ws.add_data_validation(dv_status)

    for r in range(5, 105):
        ws.cell(row=r, column=1, value=f"GAP-{r-4:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            if c == 11:  # Days Open formula
                cell.value = f'=IF(H{r}="","",IF(J{r}="Completed",DATEDIF(H{r},I{r},"d"),DATEDIF(H{r},TODAY(),"d")))'
            elif c == 12:  # Overdue formula
                cell.value = f'=IF(OR(J{r}="Completed",J{r}="",I{r}=""),"",IF(I{r}<TODAY(),"Yes","No"))'
            else:
                cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_control.add(ws.cell(row=r, column=2))
        dv_priority.add(ws.cell(row=r, column=6))
        dv_status.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "A5"

# =============================================================================
# INCIDENT TRACKER SHEET
# =============================================================================

def create_incident_tracker(ws, styles):
    """Create Incident Tracker sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "EQUIPMENT SECURITY INCIDENT TRACKER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = {
        "Incident ID": 12,
        "Date": 12,
        "Equipment Type": 18,
        "Incident Type": 15,
        "Location": 20,
        "Data Sensitivity": 18,
        "Remote Wipe": 18,
        "Hours to Report": 15,
        "Recovery Status": 18,
        "Root Cause": 30,
        "Lesson Learned": 40,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Laptop,Mobile Phone,Tablet,Storage Media,Edge Device,Other"',
        allow_blank=False
    )
    dv_incident = DataValidation(
        type="list",
        formula1='"Lost,Stolen,Damaged,Compromised,Near Miss"',
        allow_blank=False
    )
    dv_sensitivity = DataValidation(
        type="list",
        formula1='"High (PII, Financial),Medium (Internal),Low (Public),None (encrypted)"',
        allow_blank=False
    )
    dv_wipe = DataValidation(
        type="list",
        formula1='"Yes - Success,Yes - Failed,No - Not needed,No - Not possible"',
        allow_blank=False
    )
    dv_recovery = DataValidation(
        type="list",
        formula1='"Recovered,Not recovered,Insurance,Replaced"',
        allow_blank=False
    )

    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_incident)
    ws.add_data_validation(dv_sensitivity)
    ws.add_data_validation(dv_wipe)
    ws.add_data_validation(dv_recovery)

    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"INC-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=3))
        dv_incident.add(ws.cell(row=r, column=4))
        dv_sensitivity.add(ws.cell(row=r, column=6))
        dv_wipe.add(ws.cell(row=r, column=7))
        dv_recovery.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "A4"

# =============================================================================
# REMEDIATION PLAN SHEET
# =============================================================================

def create_remediation_plan(ws, styles):
    """Create Remediation Plan sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "REMEDIATION PLAN"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    columns = {
        "Action ID": 12,
        "Related Gap ID": 15,
        "Description": 50,
        "Action Type": 18,
        "Owner": 20,
        "Start Date": 12,
        "Target Date": 12,
        "Status": 15,
        "% Complete": 12,
        "Resources Required": 30,
        "Cost Estimate": 15,
        "Notes": 40,
    }

    row = 3
    for col_idx, (col_name, col_width) in enumerate(columns.items(), start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_action_type = DataValidation(
        type="list",
        formula1='"Technical,Process,Training,Documentation"',
        allow_blank=False
    )
    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Completed,Blocked"',
        allow_blank=False
    )

    ws.add_data_validation(dv_action_type)
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"ACT-{r-3:03d}").font = Font(name="Calibri", color="808080")
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_action_type.add(ws.cell(row=r, column=4))
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"

# =============================================================================
# AUDIT READINESS SHEET
# =============================================================================

def create_audit_readiness(ws, styles):
    """Create Audit Readiness checklist sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "AUDIT READINESS CHECKLIST"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Document/Evidence", "Required", "Available", "Current", "Evidence Location"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    documents = [
        "Equipment Siting Policy (ISMS-POL-A.7.8-9)",
        "Equipment Siting Assessment (S1)",
        "Off-Premises Security Assessment (S2)",
        "Equipment Removal Authorisation Forms",
        "Remote Working Policy",
        "MDM Compliance Reports",
        "Environmental Monitoring Reports",
        "Physical Security Incident Reports",
        "Equipment Loss/Theft Reports",
        "UPS and Generator Test Records",
    ]

    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_yesno)

    row = 4
    for doc in documents:
        ws.cell(row=row, column=1, value=doc)
        ws.cell(row=row, column=2, value="Yes")
        for c in range(3, 6):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            if c < 5:
                dv_yesno.add(cell)
        row += 1

    # Audit Readiness Score
    row += 1
    ws.cell(row=row, column=1, value="AUDIT READINESS SCORE:").font = Font(name="Calibri", size=12, bold=True)
    ws.cell(row=row, column=3, value='=ROUND(COUNTIF(C4:C13,"Yes")/10*100,0)&"%"').font = Font(name="Calibri", size=12, bold=True)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 50

# =============================================================================
# TREND ANALYSIS SHEET
# =============================================================================

def create_trend_analysis(ws, styles):
    """Create Trend Analysis sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "COMPLIANCE TREND ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Quarter", "A.7.8 Score", "A.7.9 Score", "Overall Score", "Open Gaps", "Incidents", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    quarters = ["Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025", "Q1 2026"]
    row = 4
    for quarter in quarters:
        ws.cell(row=row, column=1, value=quarter)
        for c in range(2, 8):
            cell = ws.cell(row=row, column=c)
            if c == 4:  # Overall formula
                cell.value = f'=IF(OR(B{row}="",C{row}=""),"",(VALUE(SUBSTITUTE(B{row},"%",""))+VALUE(SUBSTITUTE(C{row},"%","")))/2&"%")'
            else:
                cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 15

# =============================================================================
# DATA INPUT SHEET
# =============================================================================

def create_data_input(ws, styles):
    """Create Data Input sheet for source assessment data."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "DATA INPUT - SOURCE ASSESSMENT DATA"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Enter data from S1 (Equipment Siting) and S2 (Off-Premises Assets) assessments."
    ws["A2"].font = Font(name="Calibri", italic=True)

    # A.7.8 Section
    ws.merge_cells("A4:H4")
    ws["A4"] = "A.7.8 EQUIPMENT SITING DATA (from ISMS-IMP-A.7.8.S1)"
    ws["A4"].font = styles["section_header"]["font"]
    ws["A4"].fill = styles["section_header"]["fill"]

    headers = ["Sub-domain", "Score", "Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Notes"]
    header_row = 5
    row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    a78_domains = [
        "Equipment Locations",
        "Environmental Assessment",
        "Physical Security",
        "Power Infrastructure",
        "Workstation Security",
    ]

    a78_start = row
    for domain in a78_domains:
        ws.cell(row=row, column=1, value=domain)
        ws.cell(row=row, column=2, value='=IF(D{0}=0,"0%",ROUND(E{0}/D{0}*100,1)&"%")'.format(row))
        for c in range(3, 9):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # A.7.8 Total
    a78_total_row = row
    ws.cell(row=row, column=1, value="A.7.8 TOTAL").font = Font(name="Calibri", bold=True)
    ws.cell(row=row, column=2, value=f'=IF(D{row}=0,"0%",ROUND(E{row}/D{row}*100,1)&"%")').font = Font(name="Calibri", bold=True)
    ws.cell(row=row, column=3, value=f"=SUM(D{a78_start}:D{row-1})")
    ws.cell(row=row, column=4, value=f"=SUM(E{a78_start}:E{row-1})")
    ws.cell(row=row, column=5, value=f"=SUM(F{a78_start}:F{row-1})")
    ws.cell(row=row, column=6, value=f"=SUM(G{a78_start}:G{row-1})")
    ws.cell(row=row, column=7, value=f"=SUM(H{a78_start}:H{row-1})")

    row += 2

    # A.7.9 Section
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "A.7.9 OFF-PREMISES SECURITY DATA (from ISMS-IMP-A.7.9.S2)"
    ws[f"A{row}"].font = styles["section_header"]["font"]
    ws[f"A{row}"].fill = styles["section_header"]["fill"]

    row += 1
    for col_idx, header in enumerate(["Sub-domain"] + headers[1:], start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row += 1
    a79_start = row
    a79_domains = [
        "Equipment Inventory",
        "Authorisation & Tracking",
        "Protection Measures",
        "Remote Working",
        "Permanent Off-Site",
    ]

    for domain in a79_domains:
        ws.cell(row=row, column=1, value=domain)
        ws.cell(row=row, column=2, value='=IF(D{0}=0,"0%",ROUND(E{0}/D{0}*100,1)&"%")'.format(row))
        for c in range(3, 9):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # A.7.9 Total
    a79_total_row = row
    ws.cell(row=row, column=1, value="A.7.9 TOTAL").font = Font(name="Calibri", bold=True)
    ws.cell(row=row, column=2, value=f'=IF(D{row}=0,"0%",ROUND(E{row}/D{row}*100,1)&"%")').font = Font(name="Calibri", bold=True)
    ws.cell(row=row, column=3, value=f"=SUM(D{a79_start}:D{row-1})")
    ws.cell(row=row, column=4, value=f"=SUM(E{a79_start}:E{row-1})")
    ws.cell(row=row, column=5, value=f"=SUM(F{a79_start}:F{row-1})")
    ws.cell(row=row, column=6, value=f"=SUM(G{a79_start}:G{row-1})")

    row += 2

    # Overall Score
    ws.cell(row=row, column=1, value="OVERALL SCORE:").font = Font(name="Calibri", size=14, bold=True)
    ws.cell(row=row, column=2, value=f'=IF(D{a78_total_row}+D{a79_total_row}=0,"0%",ROUND((E{a78_total_row}+E{a79_total_row})/(D{a78_total_row}+D{a79_total_row})*100,1)&"%")').font = Font(name="Calibri", size=14, bold=True)

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 18

# =============================================================================
# MAIN WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    """Generate complete workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    styles = setup_styles()

    # Create all sheets
    ws = wb.create_sheet("Executive Summary", 0)
    create_executive_summary(ws, styles)

    ws = wb.create_sheet("Control Area Details")
    create_control_details(ws, styles)

    ws = wb.create_sheet("Gap Register")
    create_gap_register(ws, styles)

    ws = wb.create_sheet("Incident Tracker")
    create_incident_tracker(ws, styles)

    ws = wb.create_sheet("Remediation Plan")
    create_remediation_plan(ws, styles)

    ws = wb.create_sheet("Audit Readiness")
    create_audit_readiness(ws, styles)

    ws = wb.create_sheet("Trend Analysis")
    create_trend_analysis(ws, styles)

    ws = wb.create_sheet("Data Input")
    create_data_input(ws, styles)

    return wb

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main entry point for workbook generation."""
    try:
        logger.info("=" * 70)
        logger.info("ISMS Equipment Protection Compliance Dashboard (A.7.8-9)")
        logger.info("=" * 70)

        wb = create_workbook()
        filename = OUTPUT_FILENAME
        wb.save(filename)

        logger.info("%s SUCCESS: %s", CHECK, filename)
        logger.info("  %s 8 professional worksheets created", BULLET)
        logger.info("  %s Executive Summary with overall compliance score", BULLET)
        logger.info("  %s Control Area Details for A.7.8 and A.7.9", BULLET)
        logger.info("  %s Gap Register with aging and overdue tracking", BULLET)
        logger.info("  %s Incident Tracker for equipment security events", BULLET)
        logger.info("  %s Remediation Plan with action tracking", BULLET)
        logger.info("  %s Audit Readiness checklist", BULLET)
        logger.info("  %s Trend Analysis for quarterly reporting", BULLET)
        logger.info("  %s Data Input for source assessment integration", BULLET)
        logger.info("=" * 70)

        return 0

    except Exception as e:
        logger.error("%s ERROR: Failed to generate workbook", XMARK)
        logger.error("Error: %s", str(e))
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
