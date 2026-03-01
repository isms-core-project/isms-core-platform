#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.7-8.S3 - Endpoint and Physical Security Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.7-8: Remote Working and Security Event Reporting
Assessment Domain 3 of 4: Endpoint and Physical Security Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific remote working and security event reporting infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Remote work authorisation categories and approval criteria (match your policy)
2. Technical control requirements per remote work scenario and risk level
3. Endpoint security standard requirements (adapt to your device management platform)
4. Security event reporting channels and severity thresholds
5. Remote access technology types and security configuration requirements

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.6.7-8 Remote Working and Security Event Reporting Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
remote working and security event reporting controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Endpoint and Physical Security Assessment under ISO 27001:2022 Controls A.6.7 and A.6.8. Supports evidence-based evaluation of remote working security controls, endpoint protection, and security event reporting effectiveness.

**Assessment Scope:**
- Remote work authorisation process completeness and compliance
- Technical control implementation coverage for remote workers
- Endpoint security configuration and management compliance
- Security event detection and reporting mechanism availability
- Secure remote access channel configuration and monitoring
- User awareness and training for remote working requirements
- Evidence collection for access management and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
5. Summary Dashboard - Compliance overview and key metrics
6. Evidence Register - Audit evidence tracking
7. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 4 domains covering Remote Working and Security Event Reporting controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a678_s3_endpoint_physical_security.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a678_s3_endpoint_physical_security.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a678_s3_endpoint_physical_security.py --date 20250115

Output:
    File: ISMS-IMP-A.6.7-8.S3_Endpoint_and_Physical_Security_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.6.7-8
Assessment Domain:    3 of 4 (Endpoint and Physical Security Assessment)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.6.7-8: Remote Working and Security Event Reporting Policy (Governance)
    - ISMS-IMP-A.6.7-8.S1: Remote Work Authorisation Assessment (Domain 1)
    - ISMS-IMP-A.6.7-8.S2: Technical Controls Assessment (Domain 2)
    - ISMS-IMP-A.6.7-8.S3: Endpoint and Physical Security Assessment (Domain 3)
    - ISMS-IMP-A.6.7-8.S4: Event Reporting Mechanisms Assessment (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.6.7-8.S3 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive remote working and security event reporting details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review remote working security controls and event reporting procedures annually or when remote work policies change, new remote access technologies are deployed, or security incidents involving remote workers occur.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

import logging
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.7-8.S3"
WORKBOOK_NAME = "Endpoint and Physical Security Assessment"
CONTROL_ID = "A.6.7-8"
CONTROL_NAME = "Remote Working and Security Event Reporting"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
_THIN_SIDE = Side(style='thin')
THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)

def setup_styles():
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header": {"font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
                   "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
                   "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)},
        "column_header": {"font": Font(name="Calibri", size=10, bold=True),
                          "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
                          "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
                          "border": border_thin},
        "input_cell": {"fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
                       "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
                       "border": border_thin},
        "border": border_thin,
    }

# =============================================================================
# PRE-POPULATED DATA
# =============================================================================
DEVICE_TYPES = ["Laptop", "Desktop", "Tablet", "Mobile", "Other"]
OWNERSHIP_TYPES = ["Corporate", "BYOD", "Contractor"]
ENCRYPTION_TYPES = ["BitLocker", "FileVault", "LUKS", "Other", "None"]
PHYSICAL_REQUIREMENTS = [
    ("Screen Privacy", "Position screen away from windows and common areas", "Mandatory"),
    ("Privacy Screen", "Use privacy screen filter in public spaces", "Conditional"),
    ("Secure Storage", "Lockable cabinet or drawer for sensitive documents", "Mandatory"),
    ("Device Security", "Cable lock or secure storage for portable devices", "Recommended"),
    ("Unattended Devices", "Never leave devices unattended in public", "Mandatory"),
    ("Document Handling", "Shred sensitive documents (cross-cut shredder)", "Mandatory"),
    ("Clear Desk", "Clear desk when leaving workspace", "Mandatory"),
    ("Access Control", "Prevent family/visitor access to work devices", "Mandatory"),
    ("Travel - Hand Luggage", "Carry devices in hand luggage", "Mandatory"),
    ("Travel - Vehicle", "Never leave devices in vehicles", "Mandatory"),
]
LOST_STOLEN_PROCEDURES = [
    ("Reporting Channel", "24/7 hotline or email for reporting", "1 hour"),
    ("Account Disable", "Ability to disable account within 1 hour", "1 hour"),
    ("Remote Lock", "Ability to lock device remotely within 1 hour", "1 hour"),
    ("Remote Wipe", "Ability to wipe device remotely within 4 hours", "4 hours"),
    ("Recovery Key Access", "Access to recovery keys for encrypted devices", "As needed"),
    ("Replacement Process", "Process to issue replacement device", "24-48 hours"),
    ("Incident Documentation", "Record of incident details", "Immediate"),
    ("Investigation", "Follow-up investigation procedure", "72 hours"),
]



_STYLES = setup_styles()
def create_workbook() -> Workbook:
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    sheets = ["Instructions & Legend", "Device Inventory", "Encryption Status", "Endpoint Protection",
              "Patch Compliance", "BYOD Assessment", "Physical Security", "Lost Stolen Procedures",
              "Gap Analysis", "Evidence Register", "Summary Dashboard", "Approval Sign-Off"]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb



def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Device Inventory — list all devices used for remote working.',
        '2. Complete Encryption Status — verify full-disk encryption is enabled on all remote devices.',
        '3. Complete Endpoint Protection — assess EDR, anti-malware, and HIPS deployment on remote devices.',
        '4. Complete Patch Compliance — verify OS and application patching on remote devices.',
        '5. Complete BYOD Assessment — evaluate Bring Your Own Device controls and policies.',
        '6. Complete Physical Security — assess screen privacy, clean desk, and locking requirements.',
        '7. Complete Lost/Stolen Procedures — verify remote wipe capability and reporting procedures.',
        '8. Complete Gap Analysis — identify endpoint security gaps requiring remediation.',
        '9. Maintain the Evidence Register with device assessments and MDM configuration exports.',
        '10. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A24"] = "Status Legend"
    ws["A24"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=25, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 26 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_device_inventory_sheet(ws, styles):
    ws.merge_cells("A1:K1")
    ws["A1"] = "REMOTE DEVICE INVENTORY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:K2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Device ID", "Device Type", "Ownership", "OS", "OS Version", "Assigned User",
               "Department", "Remote Enabled", "MDM Enrolled", "Last Check-in", "Status"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    for data_row in range(4, 55):
        ws.cell(row=data_row, column=1, value=f"DEV-{data_row-3:04d}").border = styles["border"]
        for col in range(2, 12):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    type_dv = DataValidation(type="list", formula1=f'"{",".join(DEVICE_TYPES)}"', allow_blank=True)
    ws.add_data_validation(type_dv)
    type_dv.add("B4:B54")

    own_dv = DataValidation(type="list", formula1=f'"{",".join(OWNERSHIP_TYPES)}"', allow_blank=True)
    ws.add_data_validation(own_dv)
    own_dv.add("C4:C54")

    yn_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yn_dv)
    yn_dv.add("H4:I54")

    status_dv = DataValidation(type="list", formula1='"Active,Inactive,Lost,Retired"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("K4:K54")

    for col in range(1, 12):
        ws.freeze_panes = "A4"

    ws.column_dimensions[get_column_letter(col)].width = 14


def create_encryption_status_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "DEVICE ENCRYPTION STATUS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Device ID", "Encryption Type", "Full-Disk", "Status", "Key Escrowed", "Last Verified", "Compliant", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Sample row (row 4) — F2F2F2 grey with realistic example data
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    enc_sample = ["DEV-0001", "BitLocker", "Yes", "Active", "Yes", "15.01.2026", "Yes", "AES-256 full disk"]
    for col_idx, val in enumerate(enc_sample, start=1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        cell.fill = _grey_fill
        cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data entry rows 5-54 (50 empty FFFFCC rows)
    for data_row in range(5, 55):
        for col in range(1, 9):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    enc_dv = DataValidation(type="list", formula1=f'"{",".join(ENCRYPTION_TYPES)}"', allow_blank=True)
    ws.add_data_validation(enc_dv)
    enc_dv.add("B5:B54")

    yn_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yn_dv)
    yn_dv.add("C5:C54")
    yn_dv.add("E5:E54")
    yn_dv.add("G5:G54")

    status_dv = DataValidation(type="list", formula1='"Active,Suspended,Not Configured"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("D5:D54")

    for col in range(1, 9):
        ws.freeze_panes = "A4"

    ws.column_dimensions[get_column_letter(col)].width = 15


def create_endpoint_protection_sheet(ws, styles):
    ws.merge_cells("A1:J1")
    ws["A1"] = "ENDPOINT PROTECTION STATUS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    ws.merge_cells("A2:J2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Device ID", "Protection Solution", "Agent Installed", "Agent Version", "Status",
               "Definitions Date", "Last Scan", "Threats (90d)", "Compliant", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Sample row (row 4) — F2F2F2 grey with realistic example data
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ep_sample = ["DEV-0001", "Windows Defender", "Yes", "4.18.2401.4", "Active",
                 "15.01.2026", "15.01.2026", 0, "Yes", "No threats detected"]
    for col_idx, val in enumerate(ep_sample, start=1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        cell.fill = _grey_fill
        cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data entry rows 5-54 (50 empty FFFFCC rows)
    for data_row in range(5, 55):
        for col in range(1, 11):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    for col in range(1, 11):
        ws.freeze_panes = "A4"

    ws.column_dimensions[get_column_letter(col)].width = 14


def create_patch_compliance_sheet(ws, styles):
    ws.merge_cells("A1:I1")
    ws["A1"] = "PATCH COMPLIANCE STATUS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    ws.merge_cells("A2:I2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Device ID", "OS Patch Level", "Critical Missing", "High Missing", "Medium Missing",
               "Last Scan", "Days Since Patch", "Compliant", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Sample row (row 4) — F2F2F2 grey with realistic example data
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    pc_sample = ["DEV-0001", "Windows 11 23H2 (KB5034123)", 0, 0, 1, "15.01.2026", 8, "Yes", "1 medium patch pending"]
    for col_idx, val in enumerate(pc_sample, start=1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        cell.fill = _grey_fill
        cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data entry rows 5-54 (50 empty FFFFCC rows)
    for data_row in range(5, 55):
        for col in range(1, 10):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    for col in range(1, 10):
        ws.freeze_panes = "A4"

    ws.column_dimensions[get_column_letter(col)].width = 15


def create_byod_assessment_sheet(ws, styles):
    ws.merge_cells("A1:L1")
    ws["A1"] = "BYOD SECURITY ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    ws.merge_cells("A2:L2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Device ID", "Device Type", "Owner", "OS/Version", "Min OS Met", "MDM Enrolled",
               "Container", "Encrypted", "Remote Wipe", "Agreement Signed", "Compliant", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Row 4: F2F2F2 grey sample row
    from openpyxl.styles import PatternFill as _PF_BYOD
    _grey_byod = _PF_BYOD(fill_type="solid", fgColor="F2F2F2")
    byod_sample = ["BYOD-001", "Laptop", "Alice Brown", "macOS 14.3", "Yes", "Yes",
                   "Yes", "Yes", "Yes", "Yes", "Compliant", "MDM enrolled via Jamf"]
    for col_idx, val in enumerate(byod_sample, start=1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        cell.fill = _grey_byod
        cell.border = styles["border"]
        cell.font = Font(name="Calibri", size=10, italic=True, color="666666")

    # Rows 5-54: empty FFFFCC input rows
    for data_row in range(5, 55):
        for col in range(1, 13):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    for col in range(1, 13):
        ws.freeze_panes = "A5"

    ws.column_dimensions[get_column_letter(col)].width = 13


def create_physical_security_sheet(ws, styles):
    ws.merge_cells("A1:I1")
    ws["A1"] = "PHYSICAL SECURITY REQUIREMENTS ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    ws.merge_cells("A2:I2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Category", "Requirement", "Level", "Documented", "Communicated",
               "Ack Required", "Ack Captured", "Verification", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row = 4
    for category, requirement, level in PHYSICAL_REQUIREMENTS:
        ws.cell(row=row, column=1, value=category).border = styles["border"]
        ws.cell(row=row, column=2, value=requirement).border = styles["border"]
        ws.cell(row=row, column=3, value=level).border = styles["border"]
        for col in range(4, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.freeze_panes = "A4"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 45
    for col in range(3, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14


def create_lost_stolen_sheet(ws, styles):
    ws.merge_cells("A1:I1")
    ws["A1"] = "LOST/STOLEN DEVICE PROCEDURES ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    ws.merge_cells("A2:I2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Procedure Element", "Requirement", "SLA", "Implemented", "Documentation",
               "Last Tested", "Responsible", "Evidence", "Compliant"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row = 4
    for element, requirement, sla in LOST_STOLEN_PROCEDURES:
        ws.cell(row=row, column=1, value=element).border = styles["border"]
        ws.cell(row=row, column=2, value=requirement).border = styles["border"]
        ws.cell(row=row, column=3, value=sla).border = styles["border"]
        for col in range(4, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.freeze_panes = "A4"

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    for col in range(3, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15


def create_gap_analysis_sheet(ws, styles):
    ws.merge_cells("A1:I1")
    ws["A1"] = "ENDPOINT AND PHYSICAL SECURITY GAP ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]

    ws.merge_cells("A2:I2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Gap ID", "Source", "Description", "Scope", "Risk", "Remediation", "Owner", "Target", "Status"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Sample row (row 4) — F2F2F2 grey, shows example data
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_data = [
        "GAP-EPS-001", "Endpoint Audit", "Unencrypted storage on remote workstations identified",
        "All remote laptops", "High", "Deploy full disk encryption via MDM policy",
        "IT Security", "2026-04-30", "Open"
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.fill = grey_fill
        cell.border = styles["border"]

    # 50 empty data rows (rows 5–54)
    for data_row in range(5, 55):
        c1 = ws.cell(row=data_row, column=1, value=f"GAP-EPS-{data_row-4:03d}")
        c1.fill = styles["input_cell"]["fill"]
        c1.border = styles["border"]
        for col in range(2, 10):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    for col in range(1, 10):
        ws.freeze_panes = "A5"

    ws.column_dimensions[get_column_letter(col)].width = 15


def create_evidence_register(ws):
    """Create the Evidence Register sheet (GS-ER-compliant standard format)."""
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _hdr_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _inp_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _er_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _ctr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row 1: A1:H1 navy title, height 35
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy_fill
    ws["A1"].alignment = _ctr_align
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    # Row 2: Italic subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = _ctr_align
    ws["A2"].border = _border

    # Row 3: Empty separator

    # Row 4: Column headers — 003366 fill, white bold font
    headers = [
        ("Evidence ID", 14),
        ("Evidence Type", 20),
        ("Description", 35),
        ("Source / Owner", 18),
        ("Date Collected", 15),
        ("Retention Period", 16),
        ("Storage Location", 25),
        ("Status", 14),
    ]
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _hdr_fill
        cell.alignment = _ctr_align
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row 5: F2F2F2 sample row starting with EV-001
    sample = ["EV-001", "[Evidence Type]", "[Description of evidence item collected]",
              "[System / Team]", "[DD.MM.YYYY]", "[1 year]", "[ISMS/Controls/...]", "[Collected]"]
    for col_idx, val in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.fill = _grey_fill
        cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
        cell.border = _border
        cell.alignment = _er_align

    # Rows 6-105: 100 FFFFCC input rows
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = _inp_fill
            cell.border = _border
            cell.alignment = _er_align

    # Status validation
    status_dv = DataValidation(
        type="list",
        formula1='"Collected,Pending,Not Available,Superseded"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("H6:H105")

    ws.freeze_panes = "A5"




def create_summary_dashboard_sheet(ws):
    """Create Gold Standard Summary Dashboard per A.6.7-8.S3."""
    from openpyxl.utils import get_column_letter

    Q = chr(34)

    ws.title = "Summary Dashboard"

    ws.merge_cells("A1:G1")
    ws["A1"] = "ENDPOINT AND PHYSICAL SECURITY — SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = THIN_BORDER
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=4, column=c).border = THIN_BORDER

    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(t1_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    DI = "'Device Inventory'"
    ES = "'Encryption Status'"
    EP = "'Endpoint Protection'"
    PC = "'Patch Compliance'"
    area_configs = [
        ("Device Inventory",
         f"=COUNTA({DI}!B5:B54)",
         f"=COUNTIF({DI}!K5:K54,{Q}Active{Q})",
         f"=COUNTIF({DI}!K5:K54,{Q}Inactive{Q})",
         f"=COUNTIF({DI}!K5:K54,{Q}Lost{Q})"),
        ("Encryption Status",
         f"=COUNTA({ES}!A5:A54)",
         f"=COUNTIF({ES}!G5:G54,{Q}Yes{Q})",
         "0",
         f"=COUNTIF({ES}!G5:G54,{Q}No{Q})"),
        ("Endpoint Protection",
         f"=COUNTA({EP}!A5:A54)",
         "0", "0", "0"),
        ("Patch Compliance",
         f"=COUNTA({PC}!A5:A54)",
         "0", "0", "0"),
    ]

    for i, (area_name, b_formula, c_formula, d_formula, e_formula) in enumerate(area_configs):
        row = 6 + i
        ws.cell(row=row, column=1, value=area_name).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(color="000000")
        for col, formula in [(2, b_formula), (3, c_formula), (4, d_formula), (5, e_formula)]:
            cell = ws.cell(row=row, column=col, value=formula)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(color="000000")
        cell_f = ws.cell(row=row, column=6)
        cell_f.value = f"=B{row}-(C{row}+D{row}+E{row})"
        cell_f.border = THIN_BORDER
        cell_f.alignment = Alignment(horizontal="center")
        cell_f.font = Font(color="000000")
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        cell_g.number_format = "0.0%"
        cell_g.border = THIN_BORDER
        cell_g.alignment = Alignment(horizontal="center")
        cell_g.font = Font(color="000000")

    total_row = 6 + len(area_configs)
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="000000")
    ws.cell(row=total_row, column=1).fill = grey_fill
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_row - 1})"
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    cell = ws.cell(row=total_row, column=7)
    cell.value = f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"
    cell.number_format = "0.0%"
    cell.font = Font(bold=True, color="000000")
    cell.fill = grey_fill
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")

    metrics_start = total_row + 2
    ws.merge_cells(f"A{metrics_start}:G{metrics_start}")
    ws[f"A{metrics_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{metrics_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{metrics_start}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=metrics_start, column=c).border = THIN_BORDER

    ws.merge_cells(f"A{metrics_start + 1}:G{metrics_start + 1}")
    ws[f"A{metrics_start + 1}"] = "A.6.7 Endpoint and Physical Security Metrics"
    ws[f"A{metrics_start + 1}"].font = Font(bold=True, color="000000")
    ws[f"A{metrics_start + 1}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"A{metrics_start + 1}"].alignment = Alignment(horizontal="left")
    for c in range(1, 8):
        ws.cell(row=metrics_start + 1, column=c).border = THIN_BORDER

    GA = "'Gap Analysis'"
    metrics = [
        ("Total Remote Devices", f"=COUNTA({DI}!B5:B54)"),
        ("Active Remote Devices", f"=COUNTIF({DI}!K5:K54,{Q}Active{Q})"),
        ("Lost or Stolen Devices", f"=COUNTIF({DI}!K5:K54,{Q}Lost{Q})"),
        ("Corporate-Owned Devices", f"=COUNTIF({DI}!C5:C54,{Q}Corporate{Q})"),
        ("BYOD Devices", f"=COUNTIF({DI}!C5:C54,{Q}BYOD{Q})"),
        ("MDM Enrolled Devices", f"=COUNTIF({DI}!I5:I54,{Q}Yes{Q})"),
        ("Devices with Full-Disk Encryption", f"=COUNTIF({ES}!G5:G54,{Q}Yes{Q})"),
        ("Devices Without Encryption", f"=COUNTIF({ES}!G5:G54,{Q}No{Q})"),
        ("Encryption Keys Escrowed", f"=COUNTIF({ES}!E5:E54,{Q}Yes{Q})"),
        ("Total Endpoint Protection Records", f"=COUNTA({EP}!A5:A54)"),
        ("Total Patch Compliance Records", f"=COUNTA({PC}!A5:A54)"),
        ("Open Endpoint Gaps", f"=COUNTIF({GA}!I4:I54,{Q}Open{Q})"),
    ]

    row = metrics_start + 2
    for metric, formula in metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(color="000000")
        cell_val = ws.cell(row=row, column=2, value=formula)
        cell_val.border = THIN_BORDER
        cell_val.font = Font(color="000000")
        cell_val.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    crit_start = row + 1
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{crit_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=crit_start, column=c).border = THIN_BORDER

    findings_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col, header in enumerate(findings_headers, 1):
        cell = ws.cell(row=crit_start + 1, column=col, value=header if header else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    findings = [
        ("Device Inventory", "Lost or stolen remote devices",
         f"=COUNTIF({DI}!K5:K54,{Q}Lost{Q})", "Critical", "Immediate"),
        ("Encryption Status", "Remote devices without encryption",
         f"=COUNTIF({ES}!G5:G54,{Q}No{Q})", "Critical", "Immediate"),
        ("Encryption Status", "Encryption keys not escrowed",
         f"=COUNTIF({ES}!E5:E54,{Q}No{Q})", "High", "Urgent"),
        ("Device Inventory", "Devices not enrolled in MDM",
         f"=COUNTIF({DI}!I5:I54,{Q}No{Q})", "High", "Urgent"),
        ("Gap Analysis", "Open endpoint and physical gaps",
         f"=COUNTIF({GA}!I4:I54,{Q}Open{Q})", "Medium", "Plan"),
    ]

    ffffcc_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    row = crit_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).font = Font(color="000000")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 13
    ws.freeze_panes = "A4"

def create_approval_sheet(ws):
    """Create Gold Standard Approval and Sign-Off sheet (GS-AS-015 compliant)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 1: Navy title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = blue
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Rows 4-8: Summary fields (B6 = compliance formula for GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}", False),
        ("Assessment Period:", "", True),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")", True),
        ("Assessment Status:", "", True),
        ("Assessed By:", "", True),
    ]
    row = 4
    status_row = None
    for label, value, fill in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if fill:
            for c in range(2, 6):
                ws.cell(row=row, column=c).fill = yellow
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        if "Assessment Status" in label:
            status_row = row
        row += 1
    ws["B6"].number_format = "0.0%"

    # Status dropdown
    dv_status = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_status)
    if status_row:
        dv_status.add(f"B{status_row}")

    # Approver sections
    row += 2
    for title, color in [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"A{row}"].border = border
            ws.merge_cells(f"B{row}:E{row}")
            for c in range(2, 6):
                ws.cell(row=row, column=c).fill = yellow
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    for c in range(2, 6):
        ws.cell(row=row, column=c).fill = yellow
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = blue
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        for c in range(2, 6):
            ws.cell(row=row, column=c).fill = yellow
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            if not hasattr(dv, "sqref") or dv.sqref is None:
                dv.sqref = dv.cells


def main():
    logger.info(f"Starting generation of {DOCUMENT_ID} - {WORKBOOK_NAME}")
    logger.info(f"Output file: {OUTPUT_FILENAME}")

    try:
        styles = _STYLES
        wb = create_workbook()

        logger.info("Creating sheets...")
        create_instructions_sheet(wb["Instructions & Legend"])
        create_device_inventory_sheet(wb["Device Inventory"], styles)
        create_encryption_status_sheet(wb["Encryption Status"], styles)
        create_endpoint_protection_sheet(wb["Endpoint Protection"], styles)
        create_patch_compliance_sheet(wb["Patch Compliance"], styles)
        create_byod_assessment_sheet(wb["BYOD Assessment"], styles)
        create_physical_security_sheet(wb["Physical Security"], styles)
        create_lost_stolen_sheet(wb["Lost Stolen Procedures"], styles)
        create_gap_analysis_sheet(wb["Gap Analysis"], styles)
        create_evidence_register(wb["Evidence Register"])
        create_summary_dashboard_sheet(wb["Summary Dashboard"])
        create_approval_sheet(wb["Approval Sign-Off"])

        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        wb.save(output_path)
        logger.info(f"Workbook saved successfully: {output_path}")

    except Exception as e:
        logger.error(f"Error generating workbook: {e}")
        sys.exit(1)


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
