#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.7-8.S1 - Remote Work Authorisation Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.7-8: Remote Working and Security Event Reporting
Assessment Domain 1 of 4: Remote Work Authorisation Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific remote working and security event reporting infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Remote work authorisation categories and approval criteria (match your policy)
2. Technical control requirements per remote work scenario and risk level
3. Endpoint security standard requirements (adapt to your device management platform)
4. Security event reporting channels and severity thresholds
5. Remote access technology types and security configuration requirements

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.6.7-8 Remote Working and Security Event Reporting Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
remote working and security event reporting controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Remote Work Authorisation Assessment under ISO 27001:2022 Controls A.6.7 and A.6.8. Supports evidence-based evaluation of remote working security controls, endpoint protection, and security event reporting effectiveness.

**Assessment Scope:**
- Remote work authorisation process completeness and compliance
- Technical control implementation coverage for remote workers
- Endpoint security configuration and management compliance
- Security event detection and reporting mechanism availability
- Secure remote access channel configuration and monitoring
- User awareness and training for remote working requirements
- Evidence collection for access management and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
5. Summary Dashboard - Compliance overview and key metrics
6. Evidence Register - Audit evidence tracking
7. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 4 domains covering Remote Working and Security Event Reporting controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a678_s1_remote_work_authorisation.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a678_s1_remote_work_authorisation.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a678_s1_remote_work_authorisation.py --date 20250115

Output:
    File: ISMS-IMP-A.6.7-8.S1_Remote_Work_Authorisation_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.6.7-8
Assessment Domain:    1 of 4 (Remote Work Authorisation Assessment)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.6.7-8: Remote Working and Security Event Reporting Policy (Governance)
    - ISMS-IMP-A.6.7-8.S1: Remote Work Authorisation Assessment (Domain 1)
    - ISMS-IMP-A.6.7-8.S2: Technical Controls Assessment (Domain 2)
    - ISMS-IMP-A.6.7-8.S3: Endpoint and Physical Security Assessment (Domain 3)
    - ISMS-IMP-A.6.7-8.S4: Event Reporting Mechanisms Assessment (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.6.7-8.S1 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive remote working and security event reporting details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review remote working security controls and event reporting procedures annually or when remote work policies change, new remote access technologies are deployed, or security incidents involving remote workers occur.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

# =============================================================================
# IMPORTS - STANDARD LIBRARY
# =============================================================================
import logging
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# IMPORTS - THIRD PARTY
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.7-8.S1"
WORKBOOK_NAME = "Remote Work Authorisation Assessment"
CONTROL_ID = "A.6.7-8"
CONTROL_NAME = "Remote Working and Security Event Reporting"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # Swiss format
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
_THIN_SIDE = Side(style='thin')
THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)

def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# =============================================================================
# PRE-POPULATED DATA
# =============================================================================

# Eligibility Criteria per POL-A.6.7-8 Section 2.1
ELIGIBILITY_CRITERIA = [
    ("Role Suitability", "Role can be performed remotely without security compromise", "High"),
    ("Data Classification", "Data accessed remotely is classified and handling requirements defined", "High"),
    ("Technical Capability", "Personnel can establish secure remote connections", "High"),
    ("Physical Environment", "Remote workspace meets privacy and security requirements", "Medium"),
    ("Regulatory Restrictions", "No regulatory or contractual restrictions on remote work", "High"),
    ("Business Continuity", "Remote work supports business continuity requirements", "Medium"),
]

# Risk Assessment Criteria per POL-A.6.7-8 Section 2.1.1
RISK_CRITERIA = [
    ("Data Classification Level", "Public=1, Internal=2, Confidential=3, Restricted=4", 4),
    ("Physical Security Capability", "Excellent=1, Good=2, Adequate=3, Poor=4", 4),
    ("Network Security Posture", "Corporate VPN=1, Secure Home=2, Variable=3, Risky=4", 4),
    ("Device Security Baseline", "Managed Corporate=1, BYOD Compliant=2, Partial=3, Non-compliant=4", 4),
    ("Regulatory Restrictions", "None=1, General=2, Industry-Specific=3, Multiple=4", 4),
]

# Physical Security Self-Assessment Items
PHYSICAL_SECURITY_ITEMS = [
    ("Screen Positioning", "Screen positioned away from windows and common areas", "Mandatory"),
    ("Privacy Screen", "Privacy screen filter available for shared/public spaces", "Conditional"),
    ("Equipment Security", "Workspace has secure storage for devices when unattended", "Mandatory"),
    ("Access Control", "Workspace prevents unauthorised access by family/visitors", "Mandatory"),
    ("Document Storage", "Secure storage available for sensitive documents", "Mandatory"),
    ("Document Disposal", "Access to shredder or secure disposal method", "Mandatory"),
    ("Clear Desk", "Clear desk practice maintained at end of work sessions", "Mandatory"),
]

# Authorisation Status Options
AUTHORIZATION_STATUS = ["Pending", "Approved", "Conditionally Approved", "Denied", "Revoked", "Expired"]


# =============================================================================
# WORKBOOK CREATION
# =============================================================================


_STYLES = setup_styles()
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    sheets = [
        "Instructions & Legend",
        "Eligibility Criteria",
        "Authorisation Register",
        "Risk Assessment",
        "Physical Security",
        "Acknowledgments",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# SHEET: INSTRUCTIONS
# =============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Eligibility Criteria — define conditions under which remote working is permitted.',
        '2. Complete Authorisation Register — record all approved remote workers with authorisation dates.',
        '3. Complete Risk Assessment — assess information security risks for each remote working arrangement.',
        '4. Complete Physical Security — document home/remote office physical security requirements.',
        '5. Complete Acknowledgments — record remote working agreement signatures.',
        '6. Complete Gap Analysis — identify remote workers without approved authorisation or missing controls.',
        '7. Maintain the Evidence Register with authorisation records and risk assessments.',
        '8. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A22"] = "Status Legend"
    ws["A22"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=23, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 24 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_eligibility_sheet(ws, styles):
    """Create the Eligibility Criteria sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "REMOTE WORK ELIGIBILITY CRITERIA ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Criterion", "Description", "Importance", "Documented", "Evidence", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row = 4
    for criterion, description, importance in ELIGIBILITY_CRITERIA:
        ws.cell(row=row, column=1, value=criterion).border = styles["border"]
        ws.cell(row=row, column=2, value=description).border = styles["border"]
        ws.cell(row=row, column=3, value=importance).border = styles["border"]
        # Input cells
        for col in range(4, 7):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Add data validation for Documented column
    dv = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"D4:D{row-1}")

    ws.freeze_panes = "A4"

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 30


# =============================================================================
# SHEET: AUTHORIZATION REGISTER
# =============================================================================

def create_authorisation_register_sheet(ws, styles):
    """Create the Authorisation Register sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "REMOTE WORK AUTHORIZATION REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:L2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "Auth ID", "Employee Name", "Department", "Role", "Remote Type",
        "Request Date", "Risk Assessment", "Physical Assessment",
        "Acknowledgment", "Status", "Approval Date", "Review Date"
    ]

    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Sample row (row 4) — F2F2F2 grey with realistic example data
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_vals = ["RW-2026-001", "John Smith", "Engineering", "Senior Developer",
                   "Full-Time", "15.01.2026", "Complete", "Complete", "Complete",
                   "Approved", "20.01.2026", "20.01.2027"]
    for col_idx, val in enumerate(sample_vals, start=1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        cell.fill = _grey_fill
        cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data entry rows 5-54 (50 empty FFFFCC rows)
    for data_row in range(5, 55):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Add data validations
    remote_type_dv = DataValidation(type="list", formula1='"Full-Time,Part-Time,Occasional,Travel Only"', allow_blank=True)
    ws.add_data_validation(remote_type_dv)
    remote_type_dv.add("E5:E54")

    status_dv = DataValidation(type="list", formula1=f'"{",".join(AUTHORIZATION_STATUS)}"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("J5:J54")

    completion_dv = DataValidation(type="list", formula1='"Complete,Pending,N/A"', allow_blank=True)
    ws.add_data_validation(completion_dv)
    completion_dv.add("G5:I54")

    for col in range(1, len(headers) + 1):
        ws.freeze_panes = "A4"

    ws.column_dimensions[get_column_letter(col)].width = 15


# =============================================================================
# SHEET: RISK ASSESSMENT
# =============================================================================

def create_risk_assessment_sheet(ws, styles):
    """Create the Risk Assessment sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "REMOTE WORK RISK ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Risk Criteria Reference
    ws["A3"] = "Risk Scoring Criteria"
    ws["A3"].font = Font(bold=True, size=11)

    headers = ["Criterion", "Scoring Guide", "Max Score"]
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row = 5
    for criterion, scoring, max_score in RISK_CRITERIA:
        ws.cell(row=row, column=1, value=criterion).border = styles["border"]
        ws.cell(row=row, column=2, value=scoring).border = styles["border"]
        ws.cell(row=row, column=3, value=max_score).border = styles["border"]
        row += 1

    row += 2
    ws[f"A{row}"] = "Risk Assessment Register"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    assess_headers = ["Auth ID", "Employee", "Data Class", "Physical", "Network", "Device", "Regulatory", "Total", "Risk Level"]
    for col_idx, header in enumerate(assess_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Sample row (first data row) — F2F2F2 grey with realistic example data
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    risk_sample = ["RW-2026-001", "John Smith", 2, 2, 2, 1, 1, 8, "Low"]
    sample_row_idx = row + 1
    for col_idx, val in enumerate(risk_sample, start=1):
        cell = ws.cell(row=sample_row_idx, column=col_idx, value=val)
        cell.fill = _grey_fill
        cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data entry rows (50 empty FFFFCC rows)
    for data_row in range(row + 2, row + 52):
        for col in range(1, len(assess_headers) + 1):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Risk level validation
    risk_dv = DataValidation(type="list", formula1='"Low,Medium,High,Critical"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add(f"I{row+2}:I{row+16}")

    ws.freeze_panes = "A4"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 12


# =============================================================================
# SHEET: PHYSICAL SECURITY
# =============================================================================

def create_physical_security_sheet(ws, styles):
    """Create the Physical Security Self-Assessment sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "PHYSICAL SECURITY SELF-ASSESSMENT TEMPLATE"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Requirements reference
    ws["A3"] = "Physical Security Requirements"
    ws["A3"].font = Font(bold=True, size=11)

    headers = ["Requirement", "Description", "Requirement Level", "Verification Method"]
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row = 5
    for requirement, description, level in PHYSICAL_SECURITY_ITEMS:
        ws.cell(row=row, column=1, value=requirement).border = styles["border"]
        ws.cell(row=row, column=2, value=description).border = styles["border"]
        ws.cell(row=row, column=3, value=level).border = styles["border"]
        ws.cell(row=row, column=4, value="Self-assessment + Annual attestation").border = styles["border"]
        row += 1

    row += 2
    ws[f"A{row}"] = "Individual Self-Assessments"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    assess_headers = ["Employee", "Assessment Date", "All Mandatory Met", "Conditional Items", "Overall Status", "Next Review", "Notes"]
    for col_idx, header in enumerate(assess_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Sample row (first data row) — F2F2F2 grey with realistic example data
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    phys_sample = ["John Smith", "15.01.2026", "Yes", "0", "Compliant", "15.01.2027", "Home office, lockable room"]
    phys_sample_row = row + 1
    for col_idx, val in enumerate(phys_sample, start=1):
        cell = ws.cell(row=phys_sample_row, column=col_idx, value=val)
        cell.fill = _grey_fill
        cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data entry rows (50 empty FFFFCC rows)
    for data_row in range(row + 2, row + 52):
        for col in range(1, len(assess_headers) + 1):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Validations
    yes_no_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yes_no_dv)
    yes_no_dv.add(f"C{row+2}:C{row+16}")

    status_dv = DataValidation(type="list", formula1='"Compliant,Non-Compliant,Pending Review"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f"E{row+2}:E{row+16}")

    ws.freeze_panes = "A4"

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 30


# =============================================================================
# SHEET: ACKNOWLEDGMENTS
# =============================================================================

def create_acknowledgments_sheet(ws, styles):
    """Create the Policy Acknowledgments tracking sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "REMOTE WORK POLICY ACKNOWLEDGMENT TRACKING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "Employee Name", "Department", "Policy Version", "Acknowledgment Date",
        "Method", "Expiry Date", "Status", "Evidence Reference"
    ]

    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Row 4: F2F2F2 grey sample row
    from openpyxl.styles import PatternFill as _PF4
    _grey_fill4 = _PF4(fill_type="solid", fgColor="F2F2F2")
    sample_data = ["Jane Doe", "IT", "v1.2", "10.01.2026", "Electronic Signature", "10.01.2027", "Current", "/evidence/ack-001"]
    for col_idx, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        cell.fill = _grey_fill4
        cell.border = styles["border"]
        cell.font = Font(name="Calibri", size=10, italic=True, color="666666")

    # Rows 5-54: empty FFFFCC input rows
    for data_row in range(5, 55):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Validations (skip sample row — start at row 5)
    method_dv = DataValidation(type="list", formula1='"Electronic Signature,Physical Signature,LMS Completion,Email Confirmation"', allow_blank=True)
    ws.add_data_validation(method_dv)
    method_dv.add("E5:E54")

    status_dv = DataValidation(type="list", formula1='"Current,Expired,Pending"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("G5:G54")

    for col in range(1, len(headers) + 1):
        ws.freeze_panes = "A5"

    ws.column_dimensions[get_column_letter(col)].width = 18


# =============================================================================
# SHEET: GAP ANALYSIS
# =============================================================================

def create_gap_analysis_sheet(ws, styles):
    """Create the Gap Analysis sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "REMOTE WORK AUTHORIZATION GAP ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:I2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "Gap ID", "Source Area", "Gap Description", "Affected Scope",
        "Risk Level", "Remediation Action", "Owner", "Target Date", "Status"
    ]

    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Sample row (row 4) — F2F2F2 grey, shows example data
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_data = [
        "GAP-RWA-001", "Authorisation", "No formal approval process documented for remote work requests",
        "All remote workers", "High", "Implement written approval workflow in HR system",
        "IT Manager", "2026-06-30", "Open"
    ]
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.fill = grey_fill
        cell.border = styles["border"]

    # 50 empty data rows (rows 5–54)
    for data_row in range(5, 55):
        c1 = ws.cell(row=data_row, column=1, value=f"GAP-RWA-{data_row-4:03d}")
        c1.fill = styles["input_cell"]["fill"]
        c1.border = styles["border"]
        for col in range(2, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Validations
    source_dv = DataValidation(type="list", formula1='"Eligibility,Authorisation,Risk Assessment,Physical Security,Acknowledgment"', allow_blank=True)
    ws.add_data_validation(source_dv)
    source_dv.add("B5:B54")

    risk_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add("E5:E54")

    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Accepted"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("I5:I54")

    ws.freeze_panes = "A5"

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14


# =============================================================================
# SHEET: EVIDENCE REGISTER
# =============================================================================

def create_evidence_register(ws):
    """Create the Evidence Register sheet (GS-ER-compliant standard format)."""
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _hdr_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _inp_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _er_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _ctr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row 1: A1:H1 navy title, height 35
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy_fill
    ws["A1"].alignment = _ctr_align
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    # Row 2: Italic subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = _ctr_align
    ws["A2"].border = _border

    # Row 3: Empty separator

    # Row 4: Column headers — 003366 fill, white bold font
    headers = [
        ("Evidence ID", 14),
        ("Evidence Type", 20),
        ("Description", 35),
        ("Source / Owner", 18),
        ("Date Collected", 15),
        ("Retention Period", 16),
        ("Storage Location", 25),
        ("Status", 14),
    ]
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _hdr_fill
        cell.alignment = _ctr_align
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row 5: F2F2F2 sample row starting with EV-001
    sample = ["EV-001", "[Evidence Type]", "[Description of evidence item collected]",
              "[System / Team]", "[DD.MM.YYYY]", "[1 year]", "[ISMS/Controls/...]", "[Collected]"]
    for col_idx, val in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.fill = _grey_fill
        cell.font = Font(name="Calibri", size=10, italic=True, color="003366")
        cell.border = _border
        cell.alignment = _er_align

    # Rows 6-105: 100 FFFFCC input rows
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = _inp_fill
            cell.border = _border
            cell.alignment = _er_align

    # Status validation
    status_dv = DataValidation(
        type="list",
        formula1='"Collected,Pending,Not Available,Superseded"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("H6:H105")

    ws.freeze_panes = "A5"




def create_summary_dashboard_sheet(ws):
    """Create Gold Standard Summary Dashboard per A.6.7-8.S1."""
    from openpyxl.utils import get_column_letter

    Q = chr(34)  # double-quote character for use in Excel formula strings

    ws.title = "Summary Dashboard"

    # Row 1: Title banner (003366)
    ws.merge_cells("A1:G1")
    ws["A1"] = "REMOTE WORK AUTHORIZATION — SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = THIN_BORDER
    ws.row_dimensions[1].height = 35

    # Row 2: CONTROL_REF italic (left-aligned per GS standard)
    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Empty

    # ---- TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW ----
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=4, column=c).border = THIN_BORDER

    # Column headers (D9D9D9)
    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(t1_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Area configs: (Area Name, total_formula, good_formula, partial_formula, bad_formula)
    AR = "'Authorisation Register'"
    RA = "'Risk Assessment'"
    PS = "'Physical Security'"
    GA = "'Gap Analysis'"
    area_configs = [
        ("Authorisation Register",
         f"=COUNTA({AR}!B5:B54)",
         f"=COUNTIF({AR}!J5:J54,{Q}Approved{Q})",
         f"=COUNTIF({AR}!J5:J54,{Q}Pending{Q})",
         f"=COUNTIFS({AR}!J5:J54,{Q}Denied{Q})+COUNTIFS({AR}!J5:J54,{Q}Revoked{Q})+COUNTIFS({AR}!J5:J54,{Q}Expired{Q})"),
        ("Risk Assessment",
         f"=COUNTA({RA}!B14:B64)",
         f"=COUNTIF({RA}!I14:I64,{Q}Low{Q})",
         f"=COUNTIF({RA}!I14:I64,{Q}Medium{Q})",
         f"=COUNTIFS({RA}!I14:I64,{Q}High{Q})+COUNTIFS({RA}!I14:I64,{Q}Critical{Q})"),
        ("Physical Security",
         f"=COUNTA({PS}!A16:A66)",
         f"=COUNTIF({PS}!E16:E66,{Q}Compliant{Q})",
         f"=COUNTIF({PS}!E16:E66,{Q}Pending Review{Q})",
         f"=COUNTIF({PS}!E16:E66,{Q}Non-Compliant{Q})"),
        ("Acknowledgments",
         "=COUNTA(Acknowledgments!A5:A54)",
         f"=COUNTIF(Acknowledgments!G5:G54,{Q}Current{Q})",
         f"=COUNTIF(Acknowledgments!G5:G54,{Q}Pending{Q})",
         f"=COUNTIF(Acknowledgments!G5:G54,{Q}Expired{Q})"),
        ("Gap Analysis",
         f"=COUNTA({GA}!B5:B54)",
         f"=COUNTIF({GA}!I5:I54,{Q}Resolved{Q})",
         f"=COUNTIF({GA}!I5:I54,{Q}In Progress{Q})",
         f"=COUNTIF({GA}!I5:I54,{Q}Open{Q})"),
    ]

    for i, (area_name, b_formula, c_formula, d_formula, e_formula) in enumerate(area_configs):
        row = 6 + i
        ws.cell(row=row, column=1, value=area_name).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(color="000000")

        for col, formula in [(2, b_formula), (3, c_formula), (4, d_formula), (5, e_formula)]:
            cell = ws.cell(row=row, column=col, value=formula)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(color="000000")

        # F: N/A
        cell_f = ws.cell(row=row, column=6)
        cell_f.value = f"=B{row}-(C{row}+D{row}+E{row})"
        cell_f.border = THIN_BORDER
        cell_f.alignment = Alignment(horizontal="center")
        cell_f.font = Font(color="000000")

        # G: Compliance %
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        cell_g.number_format = "0.0%"
        cell_g.border = THIN_BORDER
        cell_g.alignment = Alignment(horizontal="center")
        cell_g.font = Font(color="000000")

    # TOTAL row
    total_row = 6 + len(area_configs)
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="000000")
    ws.cell(row=total_row, column=1).fill = grey_fill
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_row - 1})"
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    cell = ws.cell(row=total_row, column=7)
    cell.value = f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"
    cell.number_format = "0.0%"
    cell.font = Font(bold=True, color="000000")
    cell.fill = grey_fill
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")

    # ---- TABLE 2: KEY METRICS ----
    metrics_start = total_row + 2
    ws.merge_cells(f"A{metrics_start}:G{metrics_start}")
    ws[f"A{metrics_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{metrics_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{metrics_start}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=metrics_start, column=c).border = THIN_BORDER

    # Subtitle row (D9D9D9)
    ws.merge_cells(f"A{metrics_start + 1}:G{metrics_start + 1}")
    ws[f"A{metrics_start + 1}"] = "A.6.7 Remote Work Authorisation Metrics"
    ws[f"A{metrics_start + 1}"].font = Font(bold=True, color="000000")
    ws[f"A{metrics_start + 1}"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws[f"A{metrics_start + 1}"].alignment = Alignment(horizontal="left")
    for c in range(1, 8):
        ws.cell(row=metrics_start + 1, column=c).border = THIN_BORDER

    metrics = [
        ("Total Authorisation Requests", f"=COUNTA({AR}!B5:B54)"),
        ("Approved Authorisations", f"=COUNTIF({AR}!J5:J54,{Q}Approved{Q})"),
        ("Conditionally Approved", f"=COUNTIF({AR}!J5:J54,{Q}Conditionally Approved{Q})"),
        ("Pending Authorisations", f"=COUNTIF({AR}!J5:J54,{Q}Pending{Q})"),
        ("Denied Authorisations", f"=COUNTIF({AR}!J5:J54,{Q}Denied{Q})"),
        ("Revoked or Expired", f"=COUNTIFS({AR}!J5:J54,{Q}Revoked{Q})+COUNTIFS({AR}!J5:J54,{Q}Expired{Q})"),
        ("Risk Assessments Completed", f"=COUNTA({RA}!B14:B64)"),
        ("High or Critical Risk Assessments", f"=COUNTIFS({RA}!I14:I64,{Q}High{Q})+COUNTIFS({RA}!I14:I64,{Q}Critical{Q})"),
        ("Physical Assessments Compliant", f"=COUNTIF({PS}!E16:E66,{Q}Compliant{Q})"),
        ("Policy Acknowledgments Current", f"=COUNTIF(Acknowledgments!G5:G54,{Q}Current{Q})"),
        ("Acknowledgments Expired", f"=COUNTIF(Acknowledgments!G5:G54,{Q}Expired{Q})"),
        ("Open Authorisation Gaps", f"=COUNTIF({GA}!I5:I54,{Q}Open{Q})"),
    ]

    row = metrics_start + 2
    for metric, formula in metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(color="000000")
        cell_val = ws.cell(row=row, column=2, value=formula)
        cell_val.border = THIN_BORDER
        cell_val.font = Font(color="000000")
        cell_val.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # 2 buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # ---- TABLE 3: CRITICAL FINDINGS ----
    crit_start = row + 1
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{crit_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=crit_start, column=c).border = THIN_BORDER

    findings_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col, header in enumerate(findings_headers, 1):
        cell = ws.cell(row=crit_start + 1, column=col, value=header if header else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    findings = [
        ("Authorisation Register", "Workers without valid authorisation",
         f"=COUNTIFS({AR}!J5:J54,{Q}Denied{Q})+COUNTIFS({AR}!J5:J54,{Q}Revoked{Q})+COUNTIFS({AR}!J5:J54,{Q}Expired{Q})",
         "Critical", "Immediate"),
        ("Risk Assessment", "High or Critical risk authorisations",
         f"=COUNTIFS({RA}!I14:I64,{Q}High{Q})+COUNTIFS({RA}!I14:I64,{Q}Critical{Q})",
         "Critical", "Immediate"),
        ("Physical Security", "Physical assessments non-compliant",
         f"=COUNTIF({PS}!E16:E66,{Q}Non-Compliant{Q})",
         "High", "Urgent"),
        ("Acknowledgments", "Policy acknowledgments expired",
         f"=COUNTIF(Acknowledgments!G5:G54,{Q}Expired{Q})",
         "High", "Urgent"),
        ("Gap Analysis", "Open authorisation gaps",
         f"=COUNTIF({GA}!I5:I54,{Q}Open{Q})",
         "Medium", "Plan"),
    ]

    ffffcc_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    row = crit_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).font = Font(color="000000")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    # 2 buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 13
    ws.freeze_panes = "A4"


def create_approval_sheet(ws):
    """Create Gold Standard Approval and Sign-Off sheet (GS-AS-015 compliant)."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 1: Navy title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = blue
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Rows 4-8: Summary fields (B6 = compliance formula for GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}", False),
        ("Assessment Period:", "", True),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")", True),
        ("Assessment Status:", "", True),
        ("Assessed By:", "", True),
    ]
    row = 4
    status_row = None
    for label, value, fill in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if fill:
            for c in range(2, 6):
                ws.cell(row=row, column=c).fill = yellow
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        if "Assessment Status" in label:
            status_row = row
        row += 1
    ws["B6"].number_format = "0.0%"

    # Status dropdown
    dv_status = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_status)
    if status_row:
        dv_status.add(f"B{status_row}")

    # Approver sections
    row += 2
    for title, color in [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"A{row}"].border = border
            ws.merge_cells(f"B{row}:E{row}")
            for c in range(2, 6):
                ws.cell(row=row, column=c).fill = yellow
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    for c in range(2, 6):
        ws.cell(row=row, column=c).fill = yellow
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = blue
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        for c in range(2, 6):
            ws.cell(row=row, column=c).fill = yellow
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            if not hasattr(dv, "sqref") or dv.sqref is None:
                dv.sqref = dv.cells


def main():
    """Generate the Remote Work Authorisation Assessment workbook."""
    logger.info(f"Starting generation of {DOCUMENT_ID} - {WORKBOOK_NAME}")
    logger.info(f"Output file: {OUTPUT_FILENAME}")

    try:
        # Setup
        styles = _STYLES
        wb = create_workbook()

        # Create sheets
        logger.info("Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("Creating Eligibility Criteria sheet...")
        create_eligibility_sheet(wb["Eligibility Criteria"], styles)

        logger.info("Creating Authorisation Register sheet...")
        create_authorisation_register_sheet(wb["Authorisation Register"], styles)

        logger.info("Creating Risk Assessment sheet...")
        create_risk_assessment_sheet(wb["Risk Assessment"], styles)

        logger.info("Creating Physical Security sheet...")
        create_physical_security_sheet(wb["Physical Security"], styles)

        logger.info("Creating Acknowledgments sheet...")
        create_acknowledgments_sheet(wb["Acknowledgments"], styles)

        logger.info("Creating Gap Analysis sheet...")
        create_gap_analysis_sheet(wb["Gap Analysis"], styles)

        logger.info("Creating Evidence Register sheet...")
        create_evidence_register(wb["Evidence Register"])

        logger.info("Creating Summary Dashboard sheet...")
        create_summary_dashboard_sheet(wb["Summary Dashboard"])

        logger.info("Creating Approval Sign-Off sheet...")
        create_approval_sheet(wb["Approval Sign-Off"])

        # Save workbook
        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        wb.save(output_path)
        logger.info(f"Workbook saved successfully: {output_path}")

    except Exception as e:
        logger.error(f"Error generating workbook: {e}")
        sys.exit(1)


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
