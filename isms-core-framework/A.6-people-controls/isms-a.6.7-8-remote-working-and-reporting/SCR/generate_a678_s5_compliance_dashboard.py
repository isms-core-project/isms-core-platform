#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.7-8.S5 - Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.6.7 (Remote Working) & A.6.8 (Event Reporting)
Assessment Domain S5 of S5: Compliance Dashboard (Consolidates S1-S4)

Reference Pattern: Based on ISMS-IMP-A.6.7-8.S5 specification
================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.7-8.S5"
WORKBOOK_NAME = "Compliance Dashboard"
CONTROL_ID = "A.6.7-8"
CONTROL_NAME = "Remote Working and Security Event Reporting"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header": {"font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
                   "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
                   "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)},
        "subheader": {"font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
                      "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
                      "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)},
        "column_header": {"font": Font(name="Calibri", size=10, bold=True),
                          "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
                          "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
                          "border": border_thin},
        "input_cell": {"fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
                       "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
                       "border": border_thin},
        "compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "border": border_thin,
    }

# =============================================================================
# PRE-POPULATED DATA
# =============================================================================
ASSESSMENT_DOMAINS = [
    ("S1", "Remote Work Authorization", "ISMS-IMP-A.6.7-8.S1", "Authorization, Risk Assessment, Physical Security"),
    ("S2", "Technical Controls", "ISMS-IMP-A.6.7-8.S2", "VPN, MFA, Encryption, Logging"),
    ("S3", "Endpoint & Physical Security", "ISMS-IMP-A.6.7-8.S3", "Devices, BYOD, Lost/Stolen Procedures"),
    ("S4", "Event Reporting Mechanisms", "ISMS-IMP-A.6.7-8.S4", "Channels, Categories, Response, Culture"),
]

A67_REQUIREMENTS = [
    ("2.1", "Remote Work Authorization", "Formal authorization process established"),
    ("2.1.1", "Risk Assessment Criteria", "Risk assessment criteria documented and applied"),
    ("2.2", "Physical Security", "Physical security requirements defined and verified"),
    ("2.3", "Technical Security", "VPN, MFA, encryption controls implemented"),
    ("2.4", "Data Handling", "Data classification compliance for remote work"),
    ("2.5", "Device Security", "Corporate and BYOD device requirements met"),
    ("2.6", "Termination Procedures", "Access revocation and equipment return defined"),
]

A68_REQUIREMENTS = [
    ("3.1", "Reporting Mechanisms", "Multiple accessible reporting channels available"),
    ("3.2", "Reportable Events", "Event categories defined with examples"),
    ("3.3", "Reporting Procedures", "Clear procedures and timeframes documented"),
    ("3.4", "Non-Blame Culture", "Non-blame principles documented and communicated"),
    ("3.5", "Response and Feedback", "Response timeframes defined and measured"),
]


def create_workbook() -> Workbook:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    sheets = ["Instructions", "Executive_Summary", "Assessment_Status", "A67_Compliance",
              "A68_Compliance", "Gap_Consolidation", "Evidence_Summary", "Trend_Analysis",
              "Recommendations", "Approval_Sign_Off"]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def create_instructions_sheet(ws, styles):
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    doc_info = [("Document ID", DOCUMENT_ID),
                ("Assessment Area", "Consolidated Compliance Dashboard for A.6.7-8"),
                ("Related Policy", "ISMS-POL-A.6.7-8"),
                ("Input Documents", "S1, S2, S3, S4 Assessment Workbooks"),
                ("Version", "1.0"), ("Assessment Date", ""), ("Completed By", "")]
    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "ASSESSMENT DOMAIN OVERVIEW"
    ws[f"A{row}"].font = Font(bold=True, size=11)

    row += 1
    headers = ["Domain", "Title", "Document ID", "Scope"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for domain, title, doc_id, scope in ASSESSMENT_DOMAINS:
        ws.cell(row=row, column=1, value=domain).border = styles["border"]
        ws.cell(row=row, column=2, value=title).border = styles["border"]
        ws.cell(row=row, column=3, value=doc_id).border = styles["border"]
        ws.cell(row=row, column=4, value=scope).border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 50


def create_executive_summary_sheet(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "Executive Compliance Summary"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Overall status
    ws["A3"] = "Overall Compliance Status"
    ws["A3"].font = Font(bold=True, size=12)

    ws["A4"] = "Control A.6.7 (Remote Working)"
    ws["A4"].font = Font(bold=True)
    ws["A4"].border = styles["border"]
    ws["B4"].fill = styles["input_cell"]["fill"]
    ws["B4"].border = styles["border"]

    ws["A5"] = "Control A.6.8 (Event Reporting)"
    ws["A5"].font = Font(bold=True)
    ws["A5"].border = styles["border"]
    ws["B5"].fill = styles["input_cell"]["fill"]
    ws["B5"].border = styles["border"]

    ws["A6"] = "Combined A.6.7-8 Status"
    ws["A6"].font = Font(bold=True)
    ws["A6"].border = styles["border"]
    ws["B6"].fill = styles["input_cell"]["fill"]
    ws["B6"].border = styles["border"]

    status_dv = DataValidation(type="list", formula1='"Compliant,Partially Compliant,Non-Compliant"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("B4:B6")

    # Key Metrics
    row = 8
    ws[f"A{row}"] = "Key Metrics Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    metrics = [
        ("Remote Work Authorizations (Active)", ""),
        ("Authorization Compliance Rate", ""),
        ("Technical Controls Compliance", ""),
        ("Device Compliance Rate", ""),
        ("Reporting Channels Operational", ""),
        ("Response SLA Compliance", ""),
        ("Open Gaps (High/Critical)", ""),
        ("Total Gaps", ""),
    ]

    row += 1
    for metric, value in metrics:
        ws[f"A{row}"] = metric
        ws[f"A{row}"].border = styles["border"]
        cell = ws[f"B{row}"]
        cell.fill = styles["input_cell"]["fill"]
        cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 25


def create_assessment_status_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "Assessment Domain Status"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Domain", "Title", "Last Assessed", "Assessor", "Status",
               "Compliance %", "Open Gaps", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row = 4
    for domain, title, doc_id, scope in ASSESSMENT_DOMAINS:
        ws.cell(row=row, column=1, value=domain).border = styles["border"]
        ws.cell(row=row, column=2, value=title).border = styles["border"]
        for col in range(3, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    status_dv = DataValidation(type="list", formula1='"Complete,In Progress,Not Started,Overdue"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("E4:E7")

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16


def create_a67_compliance_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "Control A.6.7 (Remote Working) Compliance"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Section", "Requirement", "Description", "Assessed", "Status",
               "Evidence", "Gaps", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row = 4
    for section, req, desc in A67_REQUIREMENTS:
        ws.cell(row=row, column=1, value=section).border = styles["border"]
        ws.cell(row=row, column=2, value=req).border = styles["border"]
        ws.cell(row=row, column=3, value=desc).border = styles["border"]
        for col in range(4, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant,N/A"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f"E4:E{row-1}")

    yn_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yn_dv)
    yn_dv.add(f"D4:D{row-1}")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 45
    for col in range(4, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15


def create_a68_compliance_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "Control A.6.8 (Event Reporting) Compliance"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Section", "Requirement", "Description", "Assessed", "Status",
               "Evidence", "Gaps", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    row = 4
    for section, req, desc in A68_REQUIREMENTS:
        ws.cell(row=row, column=1, value=section).border = styles["border"]
        ws.cell(row=row, column=2, value=req).border = styles["border"]
        ws.cell(row=row, column=3, value=desc).border = styles["border"]
        for col in range(4, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    status_dv = DataValidation(type="list", formula1='"Compliant,Partial,Non-Compliant,N/A"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f"E4:E{row-1}")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 45
    for col in range(4, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15


def create_gap_consolidation_sheet(ws, styles):
    ws.merge_cells("A1:J1")
    ws["A1"] = "Consolidated Gap Register"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Gap ID", "Source", "Control", "Description", "Impact",
               "Risk", "Remediation", "Owner", "Target", "Status"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    for data_row in range(4, 24):
        ws.cell(row=data_row, column=1, value=f"GAP-678-{data_row-3:03d}").border = styles["border"]
        for col in range(2, 11):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    source_dv = DataValidation(type="list", formula1='"S1-Authorization,S2-Technical,S3-Endpoint,S4-Reporting"', allow_blank=True)
    ws.add_data_validation(source_dv)
    source_dv.add("B4:B23")

    control_dv = DataValidation(type="list", formula1='"A.6.7,A.6.8,Both"', allow_blank=True)
    ws.add_data_validation(control_dv)
    control_dv.add("C4:C23")

    risk_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add("F4:F23")

    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Accepted"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("J4:J23")

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14


def create_evidence_summary_sheet(ws, styles):
    ws.merge_cells("A1:G1")
    ws["A1"] = "Evidence Summary"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Evidence Category", "Source Domain", "Required", "Collected", "Location", "Last Updated", "Status"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    evidence_categories = [
        ("Remote Work Authorizations", "S1"),
        ("Risk Assessments", "S1"),
        ("Physical Security Attestations", "S1"),
        ("VPN Configuration", "S2"),
        ("MFA Enrollment Reports", "S2"),
        ("Encryption Status", "S2/S3"),
        ("Device Inventory", "S3"),
        ("Patch Compliance Reports", "S3"),
        ("Reporting Channel Tests", "S4"),
        ("Response SLA Metrics", "S4"),
        ("Training Completion", "S4"),
        ("Policy Acknowledgments", "S1/S4"),
    ]

    row = 4
    for category, source in evidence_categories:
        ws.cell(row=row, column=1, value=category).border = styles["border"]
        ws.cell(row=row, column=2, value=source).border = styles["border"]
        for col in range(3, 8):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    status_dv = DataValidation(type="list", formula1='"Complete,Partial,Missing"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f"G4:G{row-1}")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    for col in range(3, 8):
        ws.column_dimensions[get_column_letter(col)].width = 16


def create_trend_analysis_sheet(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "Compliance Trend Analysis"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Period", "A.6.7 Status", "A.6.8 Status", "Open Gaps", "Closed Gaps", "Notes"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    # Pre-populate quarterly periods
    periods = ["Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"]
    row = 4
    for period in periods:
        ws.cell(row=row, column=1, value=period).border = styles["border"]
        for col in range(2, 7):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18


def create_recommendations_sheet(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "Recommendations and Action Items"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Priority", "Area", "Recommendation", "Owner", "Target Date", "Status"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    for data_row in range(4, 14):
        for col in range(1, 7):
            cell = ws.cell(row=data_row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    priority_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(priority_dv)
    priority_dv.add("A4:A13")

    area_dv = DataValidation(type="list", formula1='"Authorization,Technical,Endpoint,Reporting,Policy,Training"', allow_blank=True)
    ws.add_data_validation(area_dv)
    area_dv.add("B4:B13")

    status_dv = DataValidation(type="list", formula1='"Planned,In Progress,Complete,Deferred"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("F4:F13")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14


def create_approval_sheet(ws, styles):
    ws.merge_cells("A1:E1")
    ws["A1"] = "Dashboard Approval and Sign-Off"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = ["Role", "Name", "Signature", "Date", "Comments"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["column_header"]["border"]

    approvers = ["IT Security Manager", "HR Director", "IT Director", "CISO", "Executive Management"]
    row = 4
    for approver in approvers:
        ws.cell(row=row, column=1, value=approver).border = styles["border"]
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 40


def main():
    logger.info(f"Starting generation of {DOCUMENT_ID} - {WORKBOOK_NAME}")
    logger.info(f"Output file: {OUTPUT_FILENAME}")

    try:
        styles = setup_styles()
        wb = create_workbook()

        logger.info("Creating sheets...")
        create_instructions_sheet(wb["Instructions"], styles)
        create_executive_summary_sheet(wb["Executive_Summary"], styles)
        create_assessment_status_sheet(wb["Assessment_Status"], styles)
        create_a67_compliance_sheet(wb["A67_Compliance"], styles)
        create_a68_compliance_sheet(wb["A68_Compliance"], styles)
        create_gap_consolidation_sheet(wb["Gap_Consolidation"], styles)
        create_evidence_summary_sheet(wb["Evidence_Summary"], styles)
        create_trend_analysis_sheet(wb["Trend_Analysis"], styles)
        create_recommendations_sheet(wb["Recommendations"], styles)
        create_approval_sheet(wb["Approval_Sign_Off"], styles)

        wb.save(OUTPUT_FILENAME)
        logger.info(f"Workbook saved successfully: {OUTPUT_FILENAME}")

    except Exception as e:
        logger.error(f"Error generating workbook: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation, constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
