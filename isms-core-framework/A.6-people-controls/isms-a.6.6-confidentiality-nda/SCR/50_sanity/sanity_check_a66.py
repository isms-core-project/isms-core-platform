#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sanity Check for A.6.6 Confidentiality/NDA Workbooks

Validates workbook structure, required sheets, and basic integrity.
"""

import logging
import sys
from pathlib import Path

from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

WORKBOOK_DIR = Path(__file__).parent.parent / "90_workbooks"

EXPECTED_WORKBOOKS = {
    "ISMS-IMP-A.6.6.1": [
        "Instructions", "Template_Registry", "Template_Versions",
        "Applicability_Matrix", "Clause_Library",
        "Evidence_Register", "Approval_SignOff"
    ],
    "ISMS-IMP-A.6.6.2": [
        "Instructions", "Active_NDAs", "Signatory_Register",
        "Expiration_Monitor", "Renewal_Tracking",
        "Evidence_Register", "Approval_SignOff"
    ],
    "ISMS-IMP-A.6.6.3": [
        "Instructions", "Periodic_Review", "Template_Adequacy",
        "Coverage_Analysis", "Compliance_Check", "Gap_Register",
        "Evidence_Register", "Approval_SignOff"
    ],
    "ISMS-IMP-A.6.6.4": [
        "Instructions", "Executive_Summary", "Coverage_Metrics",
        "Expiration_Status", "Compliance_Scorecard", "KPI_Tracker",
        "Trend_Analysis", "Approval_SignOff"
    ],
}


def check_workbook(filepath: Path, expected_sheets: list) -> tuple:
    """Check a single workbook. Returns (passed, errors)."""
    errors = []

    try:
        wb = load_workbook(filepath, data_only=False)

        for sheet_name in expected_sheets:
            if sheet_name not in wb.sheetnames:
                errors.append(f"Missing sheet: {sheet_name}")

        for sheet_name in wb.sheetnames:
            if sheet_name not in expected_sheets:
                errors.append(f"Unexpected sheet: {sheet_name}")

        if "Instructions" in wb.sheetnames:
            ws = wb["Instructions"]
            if ws["A1"].value is None:
                errors.append("Instructions sheet missing header")

        if "Approval_SignOff" in wb.sheetnames:
            ws = wb["Approval_SignOff"]
            if ws["A1"].value is None:
                errors.append("Approval_SignOff sheet missing header")

        wb.close()

    except Exception as e:
        errors.append(f"Failed to open workbook: {e}")

    return (len(errors) == 0, errors)


def main() -> int:
    logger.info("=" * 60)
    logger.info("A.6.6 Workbook Sanity Check")
    logger.info("=" * 60)

    if not WORKBOOK_DIR.exists():
        logger.error("Workbook directory not found: %s", WORKBOOK_DIR)
        return 1

    all_passed = True
    workbooks_checked = 0

    for doc_id, expected_sheets in EXPECTED_WORKBOOKS.items():
        matches = list(WORKBOOK_DIR.glob(f"{doc_id}*.xlsx"))

        if not matches:
            logger.warning("Workbook not found: %s", doc_id)
            all_passed = False
            continue

        wb_path = matches[0]
        passed, errors = check_workbook(wb_path, expected_sheets)
        workbooks_checked += 1

        if passed:
            logger.info("PASS: %s", wb_path.name)
        else:
            logger.error("FAIL: %s", wb_path.name)
            for error in errors:
                logger.error("  - %s", error)
            all_passed = False

    logger.info("=" * 60)
    logger.info("Checked %d workbooks", workbooks_checked)

    if all_passed:
        logger.info("All sanity checks PASSED")
        return 0
    else:
        logger.error("Some sanity checks FAILED")
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - UTILITY SCRIPT
# =============================================================================

