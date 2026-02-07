#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.6.4 - NDA Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.6: Confidentiality or Non-Disclosure Agreements

Assessment Domain 4 of 4: NDA Compliance Dashboard

This script generates an executive dashboard for monitoring NDA program
compliance, coverage metrics, and key performance indicators.

**Generated Workbook Structure:**
1. Instructions - Dashboard guidance
2. Executive_Summary - High-level overview
3. Coverage_Metrics - Coverage by stakeholder category
4. Expiration_Status - Expiration tracking overview
5. Compliance_Scorecard - Compliance metrics
6. KPI_Tracker - Key performance indicators
7. Trend_Analysis - Historical trends
8. Approval_SignOff - Management approval

================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

DOCUMENT_ID = "ISMS-IMP-A.6.6.4"
WORKBOOK_NAME = "NDA Compliance Dashboard"
CONTROL_ID = "A.6.6"
CONTROL_NAME = "Confidentiality or Non-Disclosure Agreements"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


def setup_styles():
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "kpi_header": {
            "font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "kpi_value": {
            "font": Font(name="Calibri", size=18, bold=True),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
        "green": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "amber": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "red": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }


def create_workbook() -> Workbook:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    for name in ["Instructions", "Executive_Summary", "Coverage_Metrics",
                 "Expiration_Status", "Compliance_Scorecard", "KPI_Tracker",
                 "Trend_Analysis", "Approval_SignOff"]:
        wb.create_sheet(title=name)
    return wb


def create_instructions_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:H3")
    ws["A3"] = CONTROL_REF
    ws["A3"].font = Font(bold=True, size=12)

    instructions = [
        "",
        "PURPOSE",
        "This dashboard provides executive-level visibility into NDA program",
        "compliance, coverage, and key performance metrics.",
        "",
        "DATA SOURCES",
        "- ISMS-IMP-A.6.6.1 (Template Registry)",
        "- ISMS-IMP-A.6.6.2 (Execution & Tracking)",
        "- ISMS-IMP-A.6.6.3 (Review & Compliance)",
        "",
        "COMPLETION STEPS",
        "1. Executive_Summary: Populate high-level metrics",
        "2. Coverage_Metrics: Update coverage by category",
        "3. Expiration_Status: Track upcoming expirations",
        "4. Compliance_Scorecard: Record compliance scores",
        "5. KPI_Tracker: Update KPI values",
        "6. Trend_Analysis: Record historical data",
        "7. Approval_SignOff: Obtain required approvals",
        "",
        "UPDATE FREQUENCY",
        "- Monthly: Update metrics and track expirations",
        "- Quarterly: Full dashboard review with CISO",
        "- Annually: Comprehensive review with Executive Management",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, text in enumerate(instructions, start=5):
        ws[f"A{i}"] = text
        if text in ["PURPOSE", "DATA SOURCES", "COMPLETION STEPS", "UPDATE FREQUENCY"]:
            ws[f"A{i}"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 80


def create_executive_summary_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "NDA Program Executive Summary"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    # KPI boxes
    ws.merge_cells("A3:B5")
    ws["A3"] = "Overall Coverage"
    ws["A3"].font = styles["kpi_header"]["font"]
    ws["A3"].fill = styles["kpi_header"]["fill"]
    ws["A3"].alignment = styles["kpi_header"]["alignment"]

    ws.merge_cells("C3:D5")
    ws["C3"].font = styles["kpi_value"]["font"]
    ws["C3"].alignment = styles["kpi_value"]["alignment"]
    ws["C3"].fill = styles["input_cell"]["fill"]

    ws.merge_cells("E3:F5")
    ws["E3"] = "Active NDAs"
    ws["E3"].font = styles["kpi_header"]["font"]
    ws["E3"].fill = styles["kpi_header"]["fill"]
    ws["E3"].alignment = styles["kpi_header"]["alignment"]

    ws.merge_cells("G3:H5")
    ws["G3"].font = styles["kpi_value"]["font"]
    ws["G3"].alignment = styles["kpi_value"]["alignment"]
    ws["G3"].fill = styles["input_cell"]["fill"]

    # Second row of KPIs
    ws.merge_cells("A7:B9")
    ws["A7"] = "Expiring (30 days)"
    ws["A7"].font = styles["kpi_header"]["font"]
    ws["A7"].fill = styles["kpi_header"]["fill"]
    ws["A7"].alignment = styles["kpi_header"]["alignment"]

    ws.merge_cells("C7:D9")
    ws["C7"].font = styles["kpi_value"]["font"]
    ws["C7"].alignment = styles["kpi_value"]["alignment"]
    ws["C7"].fill = styles["input_cell"]["fill"]

    ws.merge_cells("E7:F9")
    ws["E7"] = "Open Gaps"
    ws["E7"].font = styles["kpi_header"]["font"]
    ws["E7"].fill = styles["kpi_header"]["fill"]
    ws["E7"].alignment = styles["kpi_header"]["alignment"]

    ws.merge_cells("G7:H9")
    ws["G7"].font = styles["kpi_value"]["font"]
    ws["G7"].alignment = styles["kpi_value"]["alignment"]
    ws["G7"].fill = styles["input_cell"]["fill"]

    # Key messages section
    ws.merge_cells("A11:H11")
    ws["A11"] = "Key Messages"
    ws["A11"].font = styles["subheader"]["font"]
    ws["A11"].fill = styles["subheader"]["fill"]
    ws["A11"].alignment = styles["subheader"]["alignment"]

    headers = [("A", "Message", 50), ("B", "Priority", 12), ("C", "Owner", 20), ("D", "Action Required", 14)]
    for col, header, width in headers:
        cell = ws[f"{col}12"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.border = styles["border"]

    for row in range(13, 18):
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

    priority_val = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    priority_val.add("B13:B17")
    ws.add_data_validation(priority_val)

    yes_no = DataValidation(type="list", formula1='"Yes,No"')
    yes_no.add("D13:D17")
    ws.add_data_validation(yes_no)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12


def create_coverage_metrics_sheet(ws, styles):
    ws.merge_cells("A1:H1")
    ws["A1"] = "NDA Coverage Metrics by Stakeholder Category"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Category", 22),
        ("B", "Total_Required", 14),
        ("C", "NDA_Signed", 14),
        ("D", "Coverage_%", 12),
        ("E", "Status", 12),
        ("F", "Missing", 12),
        ("G", "Expired", 12),
        ("H", "Notes", 35),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    categories = [
        "Employees",
        "Contractors",
        "Consultants",
        "Vendors",
        "Suppliers",
        "Partners",
        "Customers (high-risk)",
        "Board Members",
        "Visitors (sensitive)",
        "TOTAL",
    ]

    for row_idx, category in enumerate(categories, start=3):
        ws[f"A{row_idx}"] = category
        if category == "TOTAL":
            ws[f"A{row_idx}"].font = Font(bold=True)
        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            if col != "A":
                ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

    status_val = DataValidation(type="list", formula1='"Green,Amber,Red"')
    status_val.add("E3:E12")
    ws.add_data_validation(status_val)

    ws.freeze_panes = "A3"


def create_expiration_status_sheet(ws, styles):
    ws.merge_cells("A1:G1")
    ws["A1"] = "NDA Expiration Status Overview"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Time_Period", 20),
        ("B", "Count", 12),
        ("C", "Counterparties", 40),
        ("D", "Renewal_Started", 14),
        ("E", "Renewal_Complete", 16),
        ("F", "At_Risk", 12),
        ("G", "Notes", 30),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    periods = [
        ("Expired (overdue)", styles["red"]["fill"]),
        ("Expiring < 30 days", styles["red"]["fill"]),
        ("Expiring 30-60 days", styles["amber"]["fill"]),
        ("Expiring 60-90 days", styles["amber"]["fill"]),
        ("Expiring 90-180 days", styles["green"]["fill"]),
        ("Beyond 180 days", styles["green"]["fill"]),
    ]

    for row_idx, (period, fill) in enumerate(periods, start=3):
        ws[f"A{row_idx}"] = period
        ws[f"A{row_idx}"].fill = fill
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            if col not in ["A"]:
                ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def create_compliance_scorecard_sheet(ws, styles):
    ws.merge_cells("A1:G1")
    ws["A1"] = "NDA Compliance Scorecard"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Compliance_Area", 35),
        ("B", "Target", 12),
        ("C", "Current", 12),
        ("D", "Status", 12),
        ("E", "Trend", 12),
        ("F", "Owner", 20),
        ("G", "Notes", 30),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    areas = [
        ("Employee NDA coverage rate", "100%"),
        ("Contractor NDA coverage rate", "100%"),
        ("Vendor NDA coverage rate", "100%"),
        ("NDAs within validity period", "100%"),
        ("NDAs with appropriate template", "100%"),
        ("NDAs securely stored", "100%"),
        ("Periodic review completion", "100%"),
        ("Gap remediation within SLA", ">95%"),
        ("Post-termination compliance", "100%"),
    ]

    for row_idx, (area, target) in enumerate(areas, start=3):
        ws[f"A{row_idx}"] = area
        ws[f"B{row_idx}"] = target
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            if col not in ["A", "B"]:
                ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

    status_val = DataValidation(type="list", formula1='"Green,Amber,Red"')
    status_val.add("D3:D15")
    ws.add_data_validation(status_val)

    trend_val = DataValidation(type="list", formula1='"Improving,Stable,Declining"')
    trend_val.add("E3:E15")
    ws.add_data_validation(trend_val)

    ws.freeze_panes = "A3"


def create_kpi_tracker_sheet(ws, styles):
    ws.merge_cells("A1:I1")
    ws["A1"] = "NDA Program Key Performance Indicators"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "KPI_ID", 10),
        ("B", "KPI_Name", 40),
        ("C", "Target", 12),
        ("D", "Current", 12),
        ("E", "Prior_Period", 14),
        ("F", "Trend", 12),
        ("G", "Status", 12),
        ("H", "Owner", 20),
        ("I", "Notes", 30),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    kpis = [
        ("KPI-01", "Overall NDA coverage rate", "100%"),
        ("KPI-02", "Renewal completion rate (before expiry)", ">95%"),
        ("KPI-03", "New joiner NDA completion (within 5 days)", "100%"),
        ("KPI-04", "Template legal review currency (<12 months)", "100%"),
        ("KPI-05", "Gap remediation SLA compliance", ">95%"),
        ("KPI-06", "Periodic review completion", "100%"),
        ("KPI-07", "Secure storage compliance", "100%"),
        ("KPI-08", "Post-termination tracking accuracy", ">98%"),
    ]

    for row_idx, (kpi_id, name, target) in enumerate(kpis, start=3):
        ws[f"A{row_idx}"] = kpi_id
        ws[f"B{row_idx}"] = name
        ws[f"C{row_idx}"] = target
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            if col not in ["A", "B", "C"]:
                ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

    trend_val = DataValidation(type="list", formula1='"Up,Down,Stable"')
    trend_val.add("F3:F15")
    ws.add_data_validation(trend_val)

    status_val = DataValidation(type="list", formula1='"On Track,At Risk,Behind"')
    status_val.add("G3:G15")
    ws.add_data_validation(status_val)

    ws.freeze_panes = "A3"


def create_trend_analysis_sheet(ws, styles):
    ws.merge_cells("A1:K1")
    ws["A1"] = "NDA Program Trend Analysis"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Period", 12),
        ("B", "Total_NDAs", 12),
        ("C", "New_Signed", 12),
        ("D", "Renewed", 12),
        ("E", "Expired", 12),
        ("F", "Coverage_%", 12),
        ("G", "Gaps_Open", 12),
        ("H", "Gaps_Closed", 12),
        ("I", "Reviews_Done", 12),
        ("J", "Status", 12),
        ("K", "Notes", 35),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    periods = ["Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025",
               "Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"]

    for row_idx, period in enumerate(periods, start=3):
        ws[f"A{row_idx}"] = period
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            if col != "A":
                ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

    status_val = DataValidation(type="list", formula1='"Green,Amber,Red"')
    status_val.add("J3:J15")
    ws.add_data_validation(status_val)

    ws.freeze_panes = "A3"


def create_approval_sheet(ws, styles):
    ws.merge_cells("A1:F1")
    ws["A1"] = "Dashboard Approval and Sign-Off"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Approval_Type", 25),
        ("B", "Approver_Role", 25),
        ("C", "Approver_Name", 25),
        ("D", "Signature", 20),
        ("E", "Date", 14),
        ("F", "Comments", 35),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    approvals = [
        ("Dashboard Prepared", "Information Security Manager"),
        ("Data Verified", "HR Manager / Contracts Manager"),
        ("Dashboard Approved", "CISO"),
        ("Presented to Management", "CISO"),
    ]

    for row_idx, (approval_type, role) in enumerate(approvals, start=3):
        ws[f"A{row_idx}"] = approval_type
        ws[f"B{row_idx}"] = role
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{col}{row_idx}"].border = styles["border"]
            if col in ["C", "D", "E", "F"]:
                ws[f"{col}{row_idx}"].fill = styles["input_cell"]["fill"]

    ws.freeze_panes = "A3"


def main() -> int:
    logger.info("=" * 60)
    logger.info("Generating %s", OUTPUT_FILENAME)
    logger.info("=" * 60)

    try:
        wb = create_workbook()
        styles = setup_styles()

        create_instructions_sheet(wb["Instructions"], styles)
        create_executive_summary_sheet(wb["Executive_Summary"], styles)
        create_coverage_metrics_sheet(wb["Coverage_Metrics"], styles)
        create_expiration_status_sheet(wb["Expiration_Status"], styles)
        create_compliance_scorecard_sheet(wb["Compliance_Scorecard"], styles)
        create_kpi_tracker_sheet(wb["KPI_Tracker"], styles)
        create_trend_analysis_sheet(wb["Trend_Analysis"], styles)
        create_approval_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)
        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        return 0

    except Exception as e:
        logger.error("FAILED: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# =============================================================================
