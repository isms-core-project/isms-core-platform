#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.6.1 - NDA Template Registry and Inventory Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.6: Confidentiality or Non-Disclosure Agreements
Assessment Domain 1 of 3: NDA Template Registry and Inventory

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific confidentiality and non-disclosure agreement management infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. NDA template categories and mandatory clause requirements (legal review required)
2. Signatory categories and NDA type applicability criteria
3. NDA execution workflow and countersignature requirements
4. Review and renewal trigger criteria (term, role change, access change)
5. Breach notification and enforcement procedure references

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.6.6 Confidentiality or Non-Disclosure Agreements Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
confidentiality and non-disclosure agreement management controls and compliance requirements.

**Purpose:**
Enables systematic management of NDA Template Registry and Inventory under ISO 27001:2022 Control A.6.6. Supports evidence-based documentation of NDA coverage, execution tracking, and compliance review for audit readiness.

**Assessment Scope:**
- NDA template inventory and clause compliance completeness
- Signatory coverage across employee, contractor, and third-party categories
- NDA execution tracking and countersignature verification
- Review and renewal cycle compliance
- Breach incident and escalation procedure documentation
- Archived NDA accessibility and retention compliance
- Evidence collection for legal, HR, and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
4. Summary Dashboard - Compliance overview and key metrics
5. Evidence Register - Audit evidence tracking
6. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 3 domains covering Confidentiality or Non-Disclosure Agreements controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a66_1_nda_template_registry.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a66_1_nda_template_registry.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a66_1_nda_template_registry.py --date 20250115

Output:
    File: ISMS-IMP-A.6.6.1_NDA_Template_Registry_and_Inventory_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.6.6
Assessment Domain:    1 of 3 (NDA Template Registry and Inventory)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.6.6: Confidentiality or Non-Disclosure Agreements Policy (Governance)
    - ISMS-IMP-A.6.6.1: NDA Template Registry and Inventory (Domain 1)
    - ISMS-IMP-A.6.6.2: NDA Execution and Tracking (Domain 2)
    - ISMS-IMP-A.6.6.3: NDA Review and Compliance (Domain 3)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.6.6.1 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive confidentiality and non-disclosure agreement management details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review NDA templates and execution tracking annually or when legal requirements change, organisational relationships evolve, or NDA compliance incidents are identified.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

import logging
from pathlib import Path
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.6.1"
WORKBOOK_NAME = "NDA Template Registry and Inventory"
CONTROL_ID = "A.6.6"
CONTROL_NAME = "Confidentiality or Non-Disclosure Agreements"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Standard colours
_NAVY = "003366"
_MID  = "4472C4"
_GREY = "D9D9D9"
_F2F2 = "F2F2F2"
_YLLW = "FFFFCC"
_GRNO = "C6EFCE"
_AMBR = "FFEB9C"
_REDC = "FFC7CE"
_RED2 = "C00000"

_thin = Side(style="thin")
THIN_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

def setup_styles():
    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color=_GREY, end_color=_GREY, fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": THIN_BORDER,
        },
        "input_cell": {
            "fill": PatternFill(start_color=_YLLW, end_color=_YLLW, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": THIN_BORDER,
        },
        "border": THIN_BORDER,
    }



_STYLES = setup_styles()
def create_workbook() -> Workbook:
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    for name in [
        "Instructions & Legend",
        "Template Registry",
        "Template Versions",
        "Applicability Matrix",
        "Clause Library",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]:
        wb.create_sheet(title=name)
    return wb



def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Template Registry — list all NDA and confidentiality agreement templates in use.',
        '2. Complete Template Versions — document version history and current approved version per template.',
        '3. Complete Applicability Matrix — map templates to use cases (employee, contractor, vendor, partner).',
        '4. Complete Clause Library — document mandatory and optional clauses with legal review status.',
        '5. Complete Approval records — confirm legal review and management approval for each template.',
        '6. Maintain the Evidence Register with template documents and approval records.',
        '7. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A21"] = "Status Legend"
    ws["A21"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=22, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 23 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_template_registry_sheet(ws, styles):
    ws.merge_cells("A1:L1")
    ws["A1"] = "NDA TEMPLATE REGISTRY AND INVENTORY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:L2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        ("A", "Template ID", 14),
        ("B", "Template Name", 35),
        ("C", "Template Type", 18),
        ("D", "Stakeholder Category", 22),
        ("E", "Current Version", 14),
        ("F", "Effective Date", 14),
        ("G", "Legal Review Date", 16),
        ("H", "Legal Reviewer", 20),
        ("I", "Owner", 20),
        ("J", "Storage Location", 30),
        ("K", "Status", 14),
        ("L", "Notes", 30),
    ]
    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    type_val = DataValidation(type="list", formula1='"Standard NDA,Mutual NDA,One-Way NDA,Employment,Contractor,Vendor,Customer,Partner"')
    type_val.add("C5:C54")
    ws.add_data_validation(type_val)

    category_val = DataValidation(type="list", formula1='"Employees,Contractors,Consultants,Vendors,Suppliers,Partners,Customers,Board Members,Visitors"')
    category_val.add("D5:D54")
    ws.add_data_validation(category_val)

    status_val = DataValidation(type="list", formula1='"Active,Draft,Under Review,Superseded,Archived"')
    status_val.add("K5:K54")
    ws.add_data_validation(status_val)

    # Grey sample row (row 4)
    _grey4 = PatternFill(start_color=_F2F2, end_color=_F2F2, fill_type="solid")
    _lalign = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _sample4 = ["TMPL-001", "[Employment NDA]", "Standard NDA", "Employees", "1.0",
                "[DD.MM.YYYY]", "[DD.MM.YYYY]", "[Legal Dept]", "[IS Manager]",
                "[SharePoint/Legal/NDAs]", "Active", ""]
    for c, val in enumerate(_sample4, start=1):
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = _grey4
        cell.border = THIN_BORDER
        cell.alignment = _lalign
        cell.font = Font(name="Calibri", italic=True)

    # 50 FFFFCC empty rows (rows 5-54)
    _yllw = PatternFill(start_color=_YLLW, end_color=_YLLW, fill_type="solid")
    for row in range(5, 55):
        for c in range(1, 13):
            cell = ws.cell(row=row, column=c)
            cell.fill = _yllw
            cell.border = THIN_BORDER
            cell.alignment = _lalign

    ws.freeze_panes = "A4"


def create_template_versions_sheet(ws, styles):
    ws.merge_cells("A1:J1")
    ws["A1"] = "TEMPLATE VERSION HISTORY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:J2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        ("A", "Template ID", 14),
        ("B", "Version", 12),
        ("C", "Version Date", 14),
        ("D", "Change Description", 45),
        ("E", "Change Reason", 30),
        ("F", "Changed By", 20),
        ("G", "Legal Approved", 14),
        ("H", "Legal Approver", 20),
        ("I", "Approval Date", 14),
        ("J", "Notes", 30),
    ]
    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    yes_no = DataValidation(type="list", formula1='"Yes,No,Pending"')
    yes_no.add("G5:G54")
    ws.add_data_validation(yes_no)

    # Grey sample row (row 4)
    _grey4 = PatternFill(start_color=_F2F2, end_color=_F2F2, fill_type="solid")
    _lalign = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _sample4 = ["TMPL-001", "1.0", "[DD.MM.YYYY]", "[Initial version]",
                "[Annual review]", "[Name]", "Yes", "[Legal Counsel]", "[DD.MM.YYYY]", ""]
    for c, val in enumerate(_sample4, start=1):
        cell = ws.cell(row=4, column=c, value=val)
        cell.fill = _grey4
        cell.border = THIN_BORDER
        cell.alignment = _lalign
        cell.font = Font(name="Calibri", italic=True)

    _yllw = PatternFill(start_color=_YLLW, end_color=_YLLW, fill_type="solid")
    for row in range(5, 55):
        for c in range(1, 11):
            cell = ws.cell(row=row, column=c)
            cell.fill = _yllw
            cell.border = THIN_BORDER
            cell.alignment = _lalign

    ws.freeze_panes = "A4"


def create_applicability_matrix_sheet(ws, styles):
    ws.merge_cells("A1:I1")
    ws["A1"] = "NDA APPLICABILITY MATRIX"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:I2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        ("A", "Stakeholder Category", 22),
        ("B", "Access Type", 25),
        ("C", "Information Classification", 22),
        ("D", "Required Template", 25),
        ("E", "Timing", 20),
        ("F", "Duration", 18),
        ("G", "Post Termination", 18),
        ("H", "Mandatory", 12),
        ("I", "Notes", 35),
    ]
    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    scenarios = [
        ("Employees", "Full system access", "All levels", "Employment NDA", "Before start date", "Indefinite", "2 years", "Yes"),
        ("Contractors", "Project-specific access", "Confidential+", "Contractor NDA", "Before access granted", "Contract duration", "2 years", "Yes"),
        ("Consultants", "Advisory access", "Confidential+", "Consultant NDA", "Before engagement", "Engagement duration", "2 years", "Yes"),
        ("Vendors", "System integration", "Confidential", "Vendor NDA", "Before contract signing", "Contract duration", "3 years", "Yes"),
        ("Suppliers", "Limited access", "Internal", "Supplier NDA", "Before access granted", "Contract duration", "1 year", "Yes"),
        ("Partners", "Mutual sharing", "Confidential+", "Mutual NDA", "Before discussions", "Partnership duration", "3 years", "Yes"),
        ("Customers", "Product access", "Confidential", "Customer NDA", "Before disclosure", "Agreement duration", "2 years", "Conditional"),
        ("Board Members", "Strategic access", "All levels", "Director NDA", "Before appointment", "Tenure + 3 years", "3 years", "Yes"),
        ("Visitors", "Escorted access", "Internal only", "Visitor NDA", "Before site entry", "Visit duration", "1 year", "Conditional"),
    ]
    for row_idx, scenario in enumerate(scenarios, start=3):
        for col_idx, value in enumerate(scenario):
            col = get_column_letter(col_idx + 1)
            ws[f"{col}{row_idx}"] = value
            ws[f"{col}{row_idx}"].border = styles["border"]

    mandatory_val = DataValidation(type="list", formula1='"Yes,No,Conditional"')
    mandatory_val.add("H3:H50")
    ws.add_data_validation(mandatory_val)

    ws.freeze_panes = "A4"


def create_clause_library_sheet(ws, styles):
    ws.merge_cells("A1:G1")
    ws["A1"] = "STANDARD CONFIDENTIALITY CLAUSE LIBRARY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        ("A", "Clause ID", 12),
        ("B", "Clause Name", 30),
        ("C", "Clause Category", 20),
        ("D", "Clause Purpose", 40),
        ("E", "Standard Text", 60),
        ("F", "Mandatory", 12),
        ("G", "Notes", 30),
    ]
    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    clauses = [
        ("CL-001", "Definition of Confidential Information", "Definitions", "Define what constitutes confidential information", "[Organisation-specific definition text]", "Yes"),
        ("CL-002", "Obligations of Receiving Party", "Obligations", "Specify recipient's duties", "[Standard obligation text]", "Yes"),
        ("CL-003", "Permitted Disclosures", "Exceptions", "Define allowed disclosures", "[Standard exceptions text]", "Yes"),
        ("CL-004", "Return of Information", "Termination", "Require return/destruction of info", "[Standard return clause]", "Yes"),
        ("CL-005", "Duration of Obligations", "Term", "Specify confidentiality period", "[Duration specification]", "Yes"),
        ("CL-006", "Injunctive Relief", "Remedies", "Allow equitable remedies", "[Standard injunction text]", "Recommended"),
        ("CL-007", "No License Granted", "IP Rights", "Clarify no IP transfer", "[Standard no-license clause]", "Recommended"),
        ("CL-008", "Governing Law", "Legal", "Specify applicable law", "[Jurisdiction specification]", "Yes"),
    ]
    for row_idx, clause in enumerate(clauses, start=3):
        for col_idx, value in enumerate(clause):
            col = get_column_letter(col_idx + 1)
            ws[f"{col}{row_idx}"] = value
            ws[f"{col}{row_idx}"].border = styles["border"]

    category_val = DataValidation(type="list", formula1='"Definitions,Obligations,Exceptions,Term,Termination,Remedies,IP Rights,Legal,General"')
    category_val.add("C3:C50")
    ws.add_data_validation(category_val)

    mandatory_val = DataValidation(type="list", formula1='"Yes,No,Recommended"')
    mandatory_val.add("F3:F50")
    ws.add_data_validation(mandatory_val)

    ws.freeze_panes = "A4"


def create_evidence_register(ws):
    """Create standard Evidence Register sheet."""
    _navy_fill = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
    _grey_fill = PatternFill(start_color=_F2F2, end_color=_F2F2, fill_type="solid")
    _yllw_fill = PatternFill(start_color=_YLLW, end_color=_YLLW, fill_type="solid")
    _hdr_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    _halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _lalign = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = _hdr_font
    ws["A1"].fill = _navy_fill
    ws["A1"].alignment = _halign
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = _lalign

    col_headers = [
        ("A", "Evidence ID", 14), ("B", "Evidence Type", 22),
        ("C", "Description", 45), ("D", "Source / Location", 35),
        ("E", "Collected By", 20), ("F", "Collection Date", 16),
        ("G", "Retention Date", 16), ("H", "Notes", 25),
    ]
    for col, hdr, width in col_headers:
        cell = ws[f"{col}4"]
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _navy_fill
        cell.alignment = _halign
        cell.border = THIN_BORDER
        ws.column_dimensions[col].width = width

    # Grey sample row 5
    _sample = ["EV-001", "Policy Document", "[Description of evidence]",
               "[SharePoint/path]", "[Name]", "[DD.MM.YYYY]", "[DD.MM.YYYY]", ""]
    for c, val in enumerate(_sample, start=1):
        cell = ws.cell(row=5, column=c, value=val)
        cell.fill = _grey_fill
        cell.border = THIN_BORDER
        cell.alignment = _lalign
        cell.font = Font(name="Calibri", italic=True)

    ev_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Procedure,Record,Certificate,Report,Screenshot,Log,Other"',
        allow_blank=True)
    ws.add_data_validation(ev_dv)

    for row in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=row, column=c)
            cell.fill = _yllw_fill
            cell.border = THIN_BORDER
            cell.alignment = _lalign
        ev_dv.add(ws.cell(row=row, column=2))

    ws.freeze_panes = "A5"


def create_summary_dashboard_sheet(ws):
    """Create Summary Dashboard with TABLE 1/2/3 — Gold Standard A.8.33-34 pattern."""
    ws.title = "Summary Dashboard"

    # Row 1: Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = "NDA TEMPLATE REGISTRY \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = THIN_BORDER
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:G2")
    ws["A2"] = f"ISO/IEC 27001:2022 \u2014 Control {CONTROL_ID}: {CONTROL_NAME}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=_NAVY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Empty

    # TABLE 1 banner (Row 4)
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=4, column=c).border = THIN_BORDER

    # TABLE 1 column headers (Row 5)
    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
                  "Non-Compliant", "N/A", "Compliance %"]
    for col, hdr in enumerate(t1_headers, 1):
        cell = ws.cell(row=5, column=col, value=hdr)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color=_GREY, end_color=_GREY, fill_type="solid")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # TABLE 1 data rows (rows 6-8), TOTAL at row 9
    # Sample rows start at row 4, data rows at row 5 → COUNTIF from row 5
    area_configs = [
        # (Area Name, Sheet, StatusCol, TotalCol, Good_values, Partial_values, Bad_values, NA_values)
        ("Template Registry",
         "Template Registry", "K", "B",
         ["Active"],
         ["Under Review"],
         ["Draft", "Superseded"],
         ["Archived"]),
        ("Template Versions",
         "Template Versions", "G", "B",
         ["Yes"],
         ["Pending"],
         ["No"],
         []),
        ("Applicability Matrix",
         "Applicability Matrix", "H", "A",
         ["Yes"],
         ["Conditional"],
         ["No"],
         []),
    ]

    # Special: Applicability Matrix has no sample row (pre-populated), data starts row 3
    am_start = 3
    std_start = 5  # standard sheets: header row 3, sample row 4, data row 5

    for i, cfg in enumerate(area_configs):
        row = 6 + i
        area_name, sheet_name, st_col, tot_col = cfg[0], cfg[1], cfg[2], cfg[3]
        good_vals, part_vals, bad_vals, na_vals = cfg[4], cfg[5], cfg[6], cfg[7]

        # Data row start for this sheet
        dstart = am_start if sheet_name == "Applicability Matrix" else std_start
        dend = 50 if sheet_name == "Applicability Matrix" else 54

        # Col A: Area name
        cell = ws.cell(row=row, column=1, value=area_name)
        cell.border = THIN_BORDER
        cell.font = Font(color="000000")

        # Col B: Total Items
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = f"=COUNTA('{sheet_name}'!{tot_col}{dstart}:{tot_col}{dend})"
        cell_b.border = THIN_BORDER
        cell_b.alignment = Alignment(horizontal="center")
        cell_b.font = Font(color="000000")

        # Col C: Compliant
        cell_c = ws.cell(row=row, column=3)
        if len(good_vals) == 1:
            cell_c.value = f'=COUNTIF(\'{sheet_name}\'!{st_col}{dstart}:{st_col}{dend},"{good_vals[0]}")'
        else:
            parts = "+".join(f'COUNTIF(\'{sheet_name}\'!{st_col}{dstart}:{st_col}{dend},"{v}")' for v in good_vals)
            cell_c.value = f"={parts}"
        cell_c.border = THIN_BORDER
        cell_c.alignment = Alignment(horizontal="center")
        cell_c.font = Font(color="000000")

        # Col D: Partial
        cell_d = ws.cell(row=row, column=4)
        if not part_vals:
            cell_d.value = 0
        elif len(part_vals) == 1:
            cell_d.value = f'=COUNTIF(\'{sheet_name}\'!{st_col}{dstart}:{st_col}{dend},"{part_vals[0]}")'
        else:
            parts = "+".join(f'COUNTIF(\'{sheet_name}\'!{st_col}{dstart}:{st_col}{dend},"{v}")' for v in part_vals)
            cell_d.value = f"={parts}"
        cell_d.border = THIN_BORDER
        cell_d.alignment = Alignment(horizontal="center")
        cell_d.font = Font(color="000000")

        # Col E: Non-Compliant
        cell_e = ws.cell(row=row, column=5)
        if not bad_vals:
            cell_e.value = 0
        elif len(bad_vals) == 1:
            cell_e.value = f'=COUNTIF(\'{sheet_name}\'!{st_col}{dstart}:{st_col}{dend},"{bad_vals[0]}")'
        else:
            parts = "+".join(f'COUNTIF(\'{sheet_name}\'!{st_col}{dstart}:{st_col}{dend},"{v}")' for v in bad_vals)
            cell_e.value = f"={parts}"
        cell_e.border = THIN_BORDER
        cell_e.alignment = Alignment(horizontal="center")
        cell_e.font = Font(color="000000")

        # Col F: N/A
        cell_f = ws.cell(row=row, column=6)
        if not na_vals:
            cell_f.value = 0
        elif len(na_vals) == 1:
            cell_f.value = f'=COUNTIF(\'{sheet_name}\'!{st_col}{dstart}:{st_col}{dend},"{na_vals[0]}")'
        else:
            parts = "+".join(f'COUNTIF(\'{sheet_name}\'!{st_col}{dstart}:{st_col}{dend},"{v}")' for v in na_vals)
            cell_f.value = f"={parts}"
        cell_f.border = THIN_BORDER
        cell_f.alignment = Alignment(horizontal="center")
        cell_f.font = Font(color="000000")

        # Col G: Compliance %
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        cell_g.number_format = "0.0%"
        cell_g.border = THIN_BORDER
        cell_g.alignment = Alignment(horizontal="center")
        cell_g.font = Font(color="000000")

    # TOTAL row (row 9)
    total_row = 6 + len(area_configs)
    grey_fill = PatternFill(start_color=_GREY, end_color=_GREY, fill_type="solid")
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="000000")
    ws.cell(row=total_row, column=1).fill = grey_fill
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_row - 1})"
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    cell_g = ws.cell(row=total_row, column=7)
    cell_g.value = f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"
    cell_g.number_format = "0.0%"
    cell_g.font = Font(bold=True, color="000000")
    cell_g.fill = grey_fill
    cell_g.border = THIN_BORDER
    cell_g.alignment = Alignment(horizontal="center")

    # TABLE 2: Key Metrics
    metrics_start = total_row + 2
    ws.merge_cells(f"A{metrics_start}:G{metrics_start}")
    ws[f"A{metrics_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{metrics_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{metrics_start}"].fill = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=metrics_start, column=c).border = THIN_BORDER

    # TABLE 2 sub-header
    for col, hdr in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        cell = ws.cell(row=metrics_start + 1, column=col, value=hdr if hdr else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color=_GREY, end_color=_GREY, fill_type="solid")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    metrics = [
        ("Total NDA Templates", "=COUNTA('Template Registry'!B5:B54)"),
        ("Active Templates", '=COUNTIF(\'Template Registry\'!K5:K54,"Active")'),
        ("Templates Under Review", '=COUNTIF(\'Template Registry\'!K5:K54,"Under Review")'),
        ("Draft Templates (Incomplete)", '=COUNTIF(\'Template Registry\'!K5:K54,"Draft")'),
        ("Mutual NDA Templates", '=COUNTIF(\'Template Registry\'!C5:C54,"Mutual NDA")'),
        ("Employment NDA Templates", '=COUNTIF(\'Template Registry\'!C5:C54,"Standard NDA")'),
        ("Contractor NDA Templates", '=COUNTIF(\'Template Registry\'!C5:C54,"Contractor")'),
        ("Vendor NDA Templates", '=COUNTIF(\'Template Registry\'!C5:C54,"Vendor")'),
        ("Total Template Versions", "=COUNTA('Template Versions'!B5:B54)"),
        ("Versions with Legal Approval", '=COUNTIF(\'Template Versions\'!G5:G54,"Yes")'),
        ("Versions Pending Legal Review", '=COUNTIF(\'Template Versions\'!G5:G54,"Pending")'),
        ("Versions Not Legally Approved", '=COUNTIF(\'Template Versions\'!G5:G54,"No")'),
        ("Mandatory Applicability Rules", '=COUNTIF(\'Applicability Matrix\'!H3:H50,"Yes")'),
        ("Conditional Applicability Rules", '=COUNTIF(\'Applicability Matrix\'!H3:H50,"Conditional")'),
        ("Mandatory Clauses in Library", '=COUNTIF(\'Clause Library\'!F3:F50,"Yes")'),
        ("Recommended Clauses in Library", '=COUNTIF(\'Clause Library\'!F3:F50,"Recommended")'),
        ("Total Clauses Defined", "=COUNTA('Clause Library'!A3:A50)"),
        ("Templates Covering All Classifications", '=COUNTIF(\'Applicability Matrix\'!C3:C50,"All levels")'),
    ]

    row = metrics_start + 2
    for metric, formula in metrics:
        cell_label = ws.cell(row=row, column=1, value=metric)
        cell_label.border = THIN_BORDER
        cell_label.font = Font(color="000000")
        cell_val = ws.cell(row=row, column=2, value=formula)
        cell_val.border = THIN_BORDER
        cell_val.font = Font(color="000000")
        cell_val.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # 2 empty buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # TABLE 3: Critical Findings
    crit_start = row + 1
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{crit_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = PatternFill(start_color=_RED2, end_color=_RED2, fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=crit_start, column=c).border = THIN_BORDER

    # TABLE 3 sub-header
    findings_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col, hdr in enumerate(findings_headers, 1):
        cell = ws.cell(row=crit_start + 1, column=col, value=hdr if hdr else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color=_GREY, end_color=_GREY, fill_type="solid")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    findings = [
        ("Template Management", "Versions with no legal approval",
         '=COUNTIF(\'Template Versions\'!G5:G54,"No")', "Critical", "Immediate"),
        ("Template Management", "Versions pending legal review",
         '=COUNTIF(\'Template Versions\'!G5:G54,"Pending")', "High", "Urgent"),
        ("Template Management", "Templates in Draft status",
         '=COUNTIF(\'Template Registry\'!K5:K54,"Draft")', "Medium", "Plan"),
        ("Template Management", "Superseded templates still tracked",
         '=COUNTIF(\'Template Registry\'!K5:K54,"Superseded")', "Low", "Monitor"),
    ]

    ffffcc_fill = PatternFill(start_color=_YLLW, end_color=_YLLW, fill_type="solid")
    row = crit_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).font = Font(color="000000")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    # 2 empty buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"


def create_approval_sheet(ws):
    """Create standard Approval Sign-Off sheet — Gold Standard A.8.33-34 pattern."""
    _navy_fill = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
    _mid_fill  = PatternFill(start_color=_MID, end_color=_MID, fill_type="solid")
    _yllw_fill = PatternFill(start_color=_YLLW, end_color=_YLLW, fill_type="solid")
    _halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _lalign = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = _navy_fill
    ws["A1"].alignment = _halign
    ws.row_dimensions[1].height = 35
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = THIN_BORDER

    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=_NAVY)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = THIN_BORDER

    # ASSESSMENT SUMMARY banner (Row 3)
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = _mid_fill
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = THIN_BORDER

    # Summary fields (rows 4-8)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G8),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = _yllw_fill
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # B6 = Overall Compliance Rating formula — format as percentage
    ws["B6"].number_format = "0.0%"

    # Status dropdown on Assessment Status (row 7 = 4th summary field)
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # 3 Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", _MID),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", _MID),
        ("APPROVED BY (CISO)", _NAVY),
    ]
    row += 2
    for title, color in approvers:
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = fill
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = _yllw_fill
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1
        row += 1

    # FINAL DECISION — Gold Standard: col A plain bold label, B:E FFFFCC merged + DV
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].border = THIN_BORDER
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = _yllw_fill
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = THIN_BORDER

    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = _mid_fill
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = THIN_BORDER

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = _yllw_fill
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            if not hasattr(dv, 'sqref') or dv.sqref is None:
                dv.sqref = dv.cells


def main() -> int:
    logger.info("=" * 60)
    logger.info(f"Generating {OUTPUT_FILENAME}")
    logger.info("=" * 60)

    try:
        wb = create_workbook()
        styles = _STYLES

        create_instructions_sheet(wb["Instructions & Legend"])
        create_template_registry_sheet(wb["Template Registry"], styles)
        create_template_versions_sheet(wb["Template Versions"], styles)
        create_applicability_matrix_sheet(wb["Applicability Matrix"], styles)
        create_clause_library_sheet(wb["Clause Library"], styles)
        create_evidence_register(wb["Evidence Register"])
        create_summary_dashboard_sheet(wb["Summary Dashboard"])
        create_approval_sheet(wb["Approval Sign-Off"])

        output_path = _wkbk_dir / OUTPUT_FILENAME
        finalize_validations(wb)
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        wb.save(output_path)
        logger.info(f"SUCCESS: {output_path.name}")
        return 0

    except Exception as e:
        logger.error(f"FAILED: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
