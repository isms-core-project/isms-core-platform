#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.4-5.S1 - Disciplinary Process Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.4: Disciplinary Process
ISO/IEC 27001:2022 Control A.6.5: Responsibilities After Termination

Assessment Domain 1 of 4: Disciplinary Process Framework

This script generates a workbook for managing the disciplinary process framework
for information security policy violations.

**Generated Workbook Structure:**
1. Instructions - Assessment guidance
2. Violation_Categories - Violation classification
3. Response_Matrix - Violation-to-consequence mapping
4. Investigation_Process - Investigation procedures
5. Case_Tracker - Disciplinary case tracking
6. Evidence_Register - Audit evidence tracking
7. Approval_SignOff - Stakeholder approval

================================================================================
"""

import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.4-5.S1"
WORKBOOK_NAME = "Disciplinary Process Assessment"
CONTROL_ID = "A.6.4-5"
CONTROL_NAME = "Disciplinary Process and Employment Exit"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "minor": {"fill": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")},
        "moderate": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "serious": {"fill": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")},
        "gross": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Violation_Categories",
        "Response_Matrix",
        "Investigation_Process",
        "Case_Tracker",
        "Evidence_Register",
        "Approval_SignOff",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================
def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet with assessment guidance."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:G3")
    ws["A3"] = CONTROL_REF
    ws["A3"].font = Font(bold=True, size=12)

    instructions = [
        "",
        "PURPOSE",
        "This workbook establishes the disciplinary process framework for information",
        "security policy violations as required by ISO 27001:2022 Control A.6.4.",
        "",
        "SCOPE",
        "- Violation categories and severity levels",
        "- Response matrix linking violations to consequences",
        "- Investigation procedures and documentation",
        "- Case tracking and trend analysis",
        "",
        "COMPLETION STEPS",
        "1. Violation_Categories: Define all security violation types",
        "2. Response_Matrix: Map violations to disciplinary responses",
        "3. Investigation_Process: Document investigation procedures",
        "4. Case_Tracker: Track disciplinary cases (ongoing)",
        "5. Evidence_Register: Link supporting evidence",
        "6. Approval_SignOff: Obtain required approvals",
        "",
        "REGULATORY ALIGNMENT",
        "- Swiss Code of Obligations (OR) Art. 328, 337",
        "- Swiss nDSG Data Protection requirements",
        "- ISO/IEC 27001:2022 Control A.6.4",
        "",
        f"Generated: {GENERATED_DATE}",
    ]

    for i, text in enumerate(instructions, start=5):
        ws[f"A{i}"] = text
        if text in ["PURPOSE", "SCOPE", "COMPLETION STEPS", "REGULATORY ALIGNMENT"]:
            ws[f"A{i}"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 80


# =============================================================================
# VIOLATION CATEGORIES SHEET
# =============================================================================
def create_violation_categories_sheet(ws, styles):
    """Create the Violation_Categories sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "VIOLATION CATEGORIES"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Category_ID", 15),
        ("B", "Category_Name", 35),
        ("C", "Severity_Level", 18),
        ("D", "Description", 50),
        ("E", "Examples", 45),
        ("F", "Investigation_Required", 18),
        ("G", "Security_Team_Involvement", 22),
        ("H", "Related_Policy", 30),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Data validations
    severity_val = DataValidation(
        type="list",
        formula1='"Minor,Moderate,Serious,Gross Misconduct"'
    )
    ws.add_data_validation(severity_val)

    yes_no_val = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(yes_no_val)

    involvement_val = DataValidation(type="list", formula1='"Yes,No,Conditional"')
    ws.add_data_validation(involvement_val)

    # Pre-populate common categories
    categories = [
        ("VIOL-001", "Unauthorized Data Sharing", "Moderate", "Sharing confidential data with unauthorized parties", "Forwarding confidential emails, sharing credentials", "Yes", "Yes", "ISMS-POL-A.5.10"),
        ("VIOL-002", "Policy Non-Compliance", "Minor", "Failure to follow documented security policies", "Not locking screen, weak passwords", "No", "Conditional", "ISMS-POL-A.5.1"),
        ("VIOL-003", "Deliberate Security Bypass", "Serious", "Intentionally circumventing security controls", "Disabling antivirus, using unauthorized software", "Yes", "Yes", "ISMS-POL-A.8.1"),
        ("VIOL-004", "Data Theft", "Gross Misconduct", "Stealing or exfiltrating organisational data", "Copying data to personal devices, selling information", "Yes", "Yes", "ISMS-POL-A.5.12"),
        ("VIOL-005", "Access Abuse", "Serious", "Using access rights beyond authorized scope", "Accessing HR records, viewing unrelated customer data", "Yes", "Yes", "ISMS-POL-A.5.15"),
        ("VIOL-006", "Acceptable Use Violation", "Minor", "Violation of acceptable use policy", "Personal use of systems, prohibited websites", "No", "No", "ISMS-POL-A.5.10"),
        ("VIOL-007", "Negligent Data Handling", "Moderate", "Careless handling of sensitive information", "Leaving documents unsecured, unencrypted transfers", "Yes", "Conditional", "ISMS-POL-A.5.14"),
        ("VIOL-008", "System Sabotage", "Gross Misconduct", "Intentional damage to systems or data", "Deleting files, introducing malware", "Yes", "Yes", "ISMS-POL-A.8.1"),
    ]

    for row_idx, category in enumerate(categories, start=4):
        for col_idx, value in enumerate(category):
            col = get_column_letter(col_idx + 1)
            cell = ws[f"{col}{row_idx}"]
            cell.value = value
            cell.border = styles["border"]

        severity_val.add(ws[f"C{row_idx}"])
        yes_no_val.add(ws[f"F{row_idx}"])
        involvement_val.add(ws[f"G{row_idx}"])

    # Additional input rows
    for row in range(len(categories) + 4, len(categories) + 24):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

        severity_val.add(ws[f"C{row}"])
        yes_no_val.add(ws[f"F{row}"])
        involvement_val.add(ws[f"G{row}"])

    ws.freeze_panes = "A4"


# =============================================================================
# RESPONSE MATRIX SHEET
# =============================================================================
def create_response_matrix_sheet(ws, styles):
    """Create the Response_Matrix sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "DISCIPLINARY RESPONSE MATRIX"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Severity_Level", 18),
        ("B", "First_Occurrence", 40),
        ("C", "Second_Occurrence", 40),
        ("D", "Third_Occurrence", 40),
        ("E", "Immediate_Actions", 40),
        ("F", "Mitigating_Factors", 45),
        ("G", "Aggravating_Factors", 45),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Pre-populate response matrix
    responses = [
        ("Minor", "Verbal warning with documented counselling", "Written warning with mandatory retraining", "Final written warning", "Document incident, remind of policy", "First offence, prompt self-reporting, inadequate training", "Repeated despite training, disregard for warnings"),
        ("Moderate", "Written warning with mandatory retraining", "Final written warning with performance plan", "Termination consideration", "Preserve evidence, restrict access if needed", "No prior violations, extenuating circumstances", "Position of trust, impact on others"),
        ("Serious", "Final written warning, possible suspension", "Termination", "N/A (typically terminated)", "Immediate access suspension, evidence preservation", "Cooperation with investigation, remediation efforts", "Attempted concealment, abuse of privileged access"),
        ("Gross Misconduct", "Immediate dismissal, potential legal action", "N/A", "N/A", "Immediate removal from premises, all access disabled", "N/A - Gross misconduct warrants immediate action", "Criminal intent, harm to third parties, financial loss"),
    ]

    for row_idx, response in enumerate(responses, start=4):
        for col_idx, value in enumerate(response):
            col = get_column_letter(col_idx + 1)
            cell = ws[f"{col}{row_idx}"]
            cell.value = value
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Apply severity-based formatting
        severity = response[0]
        if severity == "Minor":
            ws[f"A{row_idx}"].fill = styles["minor"]["fill"]
        elif severity == "Moderate":
            ws[f"A{row_idx}"].fill = styles["moderate"]["fill"]
        elif severity == "Serious":
            ws[f"A{row_idx}"].fill = styles["serious"]["fill"]
        elif severity == "Gross Misconduct":
            ws[f"A{row_idx}"].fill = styles["gross"]["fill"]

    ws.freeze_panes = "A4"


# =============================================================================
# INVESTIGATION PROCESS SHEET
# =============================================================================
def create_investigation_process_sheet(ws, styles):
    """Create the Investigation_Process sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "INVESTIGATION PROCESS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Phase", 20),
        ("B", "Activities", 55),
        ("C", "Timeline", 20),
        ("D", "Responsible_Party", 25),
        ("E", "Outputs", 40),
        ("F", "Documentation_Required", 40),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Pre-populate investigation phases
    phases = [
        ("Discovery", "Receive and log incident report; conduct preliminary assessment; determine if investigation warranted", "1-2 business days", "HR Manager, Security Team", "Incident log entry, preliminary assessment", "Incident report form, initial assessment memo"),
        ("Planning", "Define investigation scope; identify evidence sources; plan interviews; notify Legal if serious", "1 business day", "Investigation Lead", "Investigation plan", "Scope document, interview schedule"),
        ("Evidence Collection", "Collect digital evidence; conduct interviews; review documentation; preserve chain of custody", "2-5 business days", "Security Team, HR", "Evidence package", "Interview notes, evidence log, forensic reports"),
        ("Analysis", "Analyse evidence; determine facts and timeline; assess violation severity; identify factors", "2-3 business days", "Investigation Lead", "Draft findings", "Analysis workpapers, timeline reconstruction"),
        ("Decision", "Review findings with decision-makers; determine appropriate action; document rationale", "1-2 business days", "HR Director, CISO, Legal", "Decision record", "Decision memo, rationale documentation"),
        ("Action", "Communicate decision to individual; implement disciplinary action; update records", "1 business day", "HR Manager, Line Manager", "Action implementation", "Employee communication, action record"),
        ("Follow-up", "Monitor compliance; process appeals; update case status; conduct trend analysis", "Ongoing", "HR, Security", "Case closure, trend report", "Follow-up notes, closure documentation"),
    ]

    for row_idx, phase in enumerate(phases, start=4):
        for col_idx, value in enumerate(phase):
            col = get_column_letter(col_idx + 1)
            cell = ws[f"{col}{row_idx}"]
            cell.value = value
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A4"


# =============================================================================
# CASE TRACKER SHEET
# =============================================================================
def create_case_tracker_sheet(ws, styles):
    """Create the Case_Tracker sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "DISCIPLINARY CASE TRACKER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Case_ID", 18),
        ("B", "Date_Reported", 14),
        ("C", "Violation_Category", 20),
        ("D", "Status", 18),
        ("E", "Investigation_Lead", 25),
        ("F", "Date_Closed", 14),
        ("G", "Outcome", 40),
        ("H", "Notes", 40),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    # Data validations
    status_val = DataValidation(
        type="list",
        formula1='"Reported,Under Investigation,Pending Decision,Action Taken,Closed,Appealed,Appeal Resolved"'
    )
    ws.add_data_validation(status_val)

    # Pre-fill Case_ID pattern and input cells
    for row in range(4, 104):
        ws[f"A{row}"] = f"DISC-2026-{row-3:03d}"
        ws[f"A{row}"].font = Font(color="808080")

        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

        status_val.add(ws[f"D{row}"])

    ws.freeze_panes = "A4"


# =============================================================================
# EVIDENCE REGISTER SHEET
# =============================================================================
def create_evidence_register_sheet(ws, styles):
    """Create the Evidence_Register sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 25

    headers = [
        ("A", "Evidence_ID", 20),
        ("B", "Evidence_Description", 50),
        ("C", "Evidence_Type", 20),
        ("D", "Storage_Location", 50),
        ("E", "Collection_Date", 14),
        ("F", "Collected_By", 25),
    ]

    for col, header, width in headers:
        cell = ws[f"{col}3"]
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["border"]
        ws.column_dimensions[col].width = width

    type_val = DataValidation(
        type="list",
        formula1='"Policy Document,Procedure,Case Record,Training Record,Communication,Form/Template,Other"'
    )
    ws.add_data_validation(type_val)

    for row in range(4, 54):
        ws[f"A{row}"] = f"EVD-A.6.4.1-{row-3:03d}"
        ws[f"A{row}"].font = Font(color="808080")

        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]

        type_val.add(ws[f"C{row}"])

    ws.freeze_panes = "A4"


# =============================================================================
# APPROVAL SIGN-OFF SHEET
# =============================================================================
def create_approval_signoff_sheet(ws, styles):
    """Create the Approval_SignOff sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Violation Categories Defined", "=COUNTA(Violation_Categories!A4:A30)-COUNTBLANK(Violation_Categories!B4:B30)"),
        ("Response Matrix Complete", ""),
        ("Investigation Process Documented", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "" or not str(value).startswith("="):
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    reviewers = [
        ("HR Director", "", ""),
        ("Legal Counsel", "", ""),
        ("CISO", "", ""),
    ]

    row += 1
    ws[f"A{row}"] = "Role"
    ws[f"B{row}"] = "Name"
    ws[f"C{row}"] = "Signature"
    ws[f"D{row}"] = "Date"
    ws[f"E{row}"] = "Comments"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].border = styles["border"]

    row += 1
    for role, _, _ in reviewers:
        ws[f"A{row}"] = role
        ws[f"A{row}"].border = styles["border"]
        for col in ["B", "C", "D", "E"]:
            ws[f"{col}{row}"].border = styles["border"]
            ws[f"{col}{row}"].fill = styles["input_cell"]["fill"]
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 35


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/7] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/7] Creating Violation_Categories sheet...")
        create_violation_categories_sheet(wb["Violation_Categories"], styles)

        logger.info("[3/7] Creating Response_Matrix sheet...")
        create_response_matrix_sheet(wb["Response_Matrix"], styles)

        logger.info("[4/7] Creating Investigation_Process sheet...")
        create_investigation_process_sheet(wb["Investigation_Process"], styles)

        logger.info("[5/7] Creating Case_Tracker sheet...")
        create_case_tracker_sheet(wb["Case_Tracker"], styles)

        logger.info("[6/7] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[7/7] Creating Approval_SignOff sheet...")
        create_approval_signoff_sheet(wb["Approval_SignOff"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        logger.error("Install with: pip install openpyxl")
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE
# QA_TOOL: Claude Code
# CHANGES: Initial creation per ISMS-IMP-A.6.4-5.S1 specification
# =============================================================================
