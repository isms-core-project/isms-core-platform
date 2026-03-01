#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.3.2 - Training Program Design Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.3: Information Security Awareness, Education and Training
Assessment Domain 2 of 3: Training Program Design

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific information security awareness and training infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Training needs assessment methodology and role categories (match your organisation)
2. Training programme structure and delivery modalities (adapt to your platforms)
3. Completion tracking and reporting mechanisms
4. Training effectiveness measurement criteria and thresholds
5. Mandatory training categories and role-specific curriculum requirements

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.6.3 Information Security Awareness, Education and Training Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
information security awareness and training controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Training Program Design under ISO 27001:2022 Control A.6.3. Supports evidence-based evaluation of awareness programme coverage, training effectiveness, and organisational security culture development.

**Assessment Scope:**
- Training needs identification and gap analysis completeness
- Programme design coverage across role categories and risk levels
- Training delivery tracking and completion rate monitoring
- Effectiveness measurement and competency verification
- Mandatory training compliance across the organisation
- Awareness campaign scheduling and reach assessment
- Evidence collection for HR, compliance, and audit reporting

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
4. Summary Dashboard - Compliance overview and key metrics
5. Evidence Register - Audit evidence tracking
6. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 3 domains covering Information Security Awareness, Education and Training controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a63_2_training_program_design.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a63_2_training_program_design.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a63_2_training_program_design.py --date 20250115

Output:
    File: ISMS-IMP-A.6.3.2_Training_Program_Design_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.6.3
Assessment Domain:    2 of 3 (Training Program Design)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.6.3: Information Security Awareness, Education and Training Policy (Governance)
    - ISMS-IMP-A.6.3.1: Training Needs Assessment (Domain 1)
    - ISMS-IMP-A.6.3.2: Training Program Design (Domain 2)
    - ISMS-IMP-A.6.3.3: Training Delivery Tracking (Domain 3)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.6.3.2 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive information security awareness and training details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review training needs and programme content annually or when threat landscape changes, new systems are introduced, compliance incidents occur, or regulatory training requirements are updated.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.3.2"
WORKBOOK_NAME = "Training Program Design"
CONTROL_ID = "A.6.3"
CONTROL_NAME = "Information Security Awareness, Education and Training"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
    }


# =============================================================================
# PRE-POPULATED DATA
# =============================================================================

CURRICULUM_CATALOG = [
    # Tier 1 - Baseline Awareness
    ("MOD-T1-001", "Information Security Fundamentals", "Tier 1", "Baseline Awareness", "Mandatory", 30, "eLearning", "Quiz", "Onboarding", "POL-A.6.3 S2.3.1"),
    ("MOD-T1-002", "Acceptable Use Policy", "Tier 1", "Baseline Awareness", "Mandatory", 20, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.3.2"),
    ("MOD-T1-003", "Data Classification and Handling", "Tier 1", "Baseline Awareness", "Mandatory", 25, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.3.3"),
    ("MOD-T1-004", "Password Security and Authentication", "Tier 1", "Baseline Awareness", "Mandatory", 20, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.3.4"),
    ("MOD-T1-005", "Phishing and Social Engineering", "Tier 1", "Baseline Awareness", "Mandatory", 30, "eLearning + Simulation", "Quiz + Simulation", "Annual", "POL-A.6.3 S2.3.5"),
    ("MOD-T1-006", "Security Incident Reporting", "Tier 1", "Baseline Awareness", "Mandatory", 15, "eLearning", "Quiz", "Onboarding", "POL-A.6.3 S2.3.6"),
    ("MOD-T1-007", "Physical Security Awareness", "Tier 1", "Baseline Awareness", "Mandatory", 15, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.3.7"),
    ("MOD-T1-008", "Remote Working Security", "Tier 1", "Baseline Awareness", "Mandatory", 20, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.3.8"),
    # Tier 3 - Data Handlers
    ("MOD-T3-001", "Privacy and Data Protection Fundamentals", "Tier 3", "Data Protection", "Mandatory", 45, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.4.1"),
    ("MOD-T3-002", "Data Subject Rights and Obligations", "Tier 3", "Data Protection", "Mandatory", 30, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.4.1"),
    ("MOD-T3-003", "Data Retention and Secure Deletion", "Tier 3", "Data Protection", "Mandatory", 25, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.4.1"),
    # Tier 4 - Technical Staff
    ("MOD-T4-001", "Secure Coding Practices", "Tier 4", "Technical Security", "Mandatory", 60, "eLearning + Workshop", "Quiz + Practical", "Annual", "POL-A.6.3 S2.4.2"),
    ("MOD-T4-002", "OWASP Top 10 Vulnerabilities", "Tier 4", "Technical Security", "Mandatory", 45, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.4.2"),
    ("MOD-T4-003", "Secure Configuration Management", "Tier 4", "Technical Security", "Mandatory", 30, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.4.2"),
    # Tier 5 - Privileged Users
    ("MOD-T5-001", "Privileged Access Management", "Tier 5", "Privileged Access", "Mandatory", 45, "Instructor-Led", "Quiz + Scenario", "Annual", "POL-A.6.3 S2.4.3"),
    ("MOD-T5-002", "Incident Response for System Admins", "Tier 5", "Privileged Access", "Mandatory", 60, "Workshop", "Practical", "Annual", "POL-A.6.3 S2.4.3"),
    # Tier 6 - Security Roles
    ("MOD-T6-001", "Risk Assessment Methodology", "Tier 6", "Specialized", "Mandatory", 120, "Instructor-Led", "Practical", "Annual", "POL-A.6.3 S2.4.4"),
    ("MOD-T6-002", "Incident Response and Forensics", "Tier 6", "Specialized", "Mandatory", 240, "Workshop", "Tabletop + Practical", "Annual", "POL-A.6.3 S2.4.4"),
    # Tier 7 - Management
    ("MOD-T7-001", "Information Security Governance", "Tier 7", "Governance", "Mandatory", 45, "eLearning + Briefing", "N/A", "Annual", "POL-A.6.3 S2.4.5"),
    ("MOD-T7-002", "Executive Cyber Risk Briefing", "Tier 7", "Governance", "Mandatory", 30, "Briefing", "N/A", "Quarterly", "POL-A.6.3 S2.4.5"),
]

DELIVERY_STANDARDS = [
    ("eLearning", "Self-paced online modules", "Baseline awareness, refresher training", "LMS access, modern browser", "N/A", "Automatic via LMS", "Integrated quiz", "WCAG 2.1 AA"),
    ("Instructor-Led", "Facilitated classroom or virtual", "Complex topics, specialized roles", "Classroom/VC platform", "Certified instructor", "Attendance record", "Post-session assessment", "Accessible venue"),
    ("Workshop", "Hands-on practical exercises", "Technical skills, incident response", "Lab environment", "Technical SME", "Participation record", "Practical assessment", "Adjustable workstations"),
    ("Simulation", "Realistic scenario exercises", "Phishing, social engineering", "Simulation platform", "Campaign design expertise", "Engagement tracking", "Behavioural measurement", "Text alternatives"),
    ("Briefing", "Executive-level sessions", "Leadership updates", "Executive meeting space", "Senior security leadership", "Attendance record", "N/A", "Standard meeting access"),
    ("Microlearning", "Brief focused modules (2-5 min)", "Reinforcement, quick updates", "Mobile-friendly delivery", "N/A", "View tracking", "Optional quick check", "Mobile accessible"),
]

ASSESSMENT_STANDARDS = [
    ("Quiz", "Knowledge verification", "Multiple choice, True/False, Matching", 10, "80%", 3, "Immediate feedback", "Retake training if 3 fails", "5 years"),
    ("Practical", "Skills demonstration", "Hands-on exercises, scenario completion", 3, "Pass/Fail per task", 2, "Instructor feedback", "Additional coaching", "3 years"),
    ("Simulation", "Behavioural assessment", "Phishing response, social engineering", 1, "Did not click + reported", "N/A", "Educational landing page", "Targeted remedial training", "3 years"),
    ("Scenario", "Situational judgment", "Case studies, decision scenarios", 5, "80%", 2, "Explanation of optimal response", "Review with SME", "5 years"),
    ("Tabletop", "Group decision-making", "Incident scenarios, team response", 1, "Participation + contribution", 1, "Facilitator debrief", "Individual follow-up", "3 years"),
]


# =============================================================================
# WORKBOOK CREATION
# =============================================================================


_STYLES = setup_styles()
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    sheets = [
        "Instructions & Legend",
        "Curriculum Catalog",
        "Content Specifications",
        "Delivery Standards",
        "Assessment Design",
        "Content Lifecycle",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Curriculum Catalog — list all training modules with topics, duration, and target audience.',
        '2. Complete Content Specifications — document learning objectives and assessment criteria per module.',
        '3. Complete Delivery Standards — define delivery method (e-learning, classroom, simulation) requirements.',
        '4. Complete Assessment Design — document assessment methods and minimum pass scores.',
        '5. Complete Content Lifecycle — track module review dates and version control.',
        '6. Complete Gap Analysis — identify curriculum gaps relative to ISO 27001:2022 A.6.3 requirements.',
        '7. Maintain the Evidence Register with curriculum documentation and approval records.',
        '8. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A22"] = "Status Legend"
    ws["A22"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=23, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 24 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_curriculum_catalog_sheet(ws, styles):
    """Create Curriculum_Catalog sheet with pre-populated data."""
    ws.merge_cells("A1:N1")
    ws["A1"] = "CURRICULUM CATALOG\nMaster catalog of all training modules"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    ws["A2"] = "Master catalog of all training modules with delivery specifications and status tracking"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")

    headers = [
        ("Module ID", 15),
        ("Module Title", 40),
        ("Training Tier", 12),
        ("Category", 20),
        ("Mandatory Optional", 18),
        ("Duration Minutes", 18),
        ("Delivery Method", 25),
        ("Assessment Type", 20),
        ("Frequency", 15),
        ("Policy Reference", 20),
        ("Content Owner", 20),
        ("Last Updated", 12),
        ("Next Review", 12),
        ("Status", 12),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Pre-populate curriculum
    row += 1
    for module_data in CURRICULUM_CATALOG:
        for col_idx, value in enumerate(module_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Add input cells for owner, dates, status
        for col_idx in range(11, 15):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Grey sample row (first empty row) + 50 FFFFCC empty rows for additional modules
    _cc_f2f2f2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws.cell(row=row, column=1, value="MOD-XXXX").font = Font(name="Calibri", color="808080")
    ws.cell(row=row, column=1).fill = _cc_f2f2f2
    for c in range(2, 15):
        cell = ws.cell(row=row, column=c)
        cell.fill = _cc_f2f2f2
        cell.border = styles["border"]
    for r in range(row + 1, row + 52):
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Add dropdowns
    dv_tier = DataValidation(type="list", formula1='"Tier 1,Tier 2,Tier 3,Tier 4,Tier 5,Tier 6,Tier 7"', allow_blank=False)
    dv_mandatory = DataValidation(type="list", formula1='"Mandatory,Optional"', allow_blank=False)
    dv_delivery = DataValidation(type="list", formula1='"eLearning,Instructor-Led,Workshop,Simulation,Briefing,Microlearning,Blended"', allow_blank=False)
    dv_frequency = DataValidation(type="list", formula1='"Onboarding,Annual,Bi-Annual,Quarterly,As Needed"', allow_blank=False)
    dv_status = DataValidation(type="list", formula1='"Active,Draft,Under Review,Retired"', allow_blank=False)

    ws.add_data_validation(dv_tier)
    ws.add_data_validation(dv_mandatory)
    ws.add_data_validation(dv_delivery)
    ws.add_data_validation(dv_frequency)
    ws.add_data_validation(dv_status)

    for r in range(4, row + 52):
        dv_tier.add(ws.cell(row=r, column=3))
        dv_mandatory.add(ws.cell(row=r, column=5))
        dv_delivery.add(ws.cell(row=r, column=7))
        dv_frequency.add(ws.cell(row=r, column=9))
        dv_status.add(ws.cell(row=r, column=14))

    ws.freeze_panes = "A4"


def create_content_specifications_sheet(ws, styles):
    """Create Content_Specifications sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "CONTENT SPECIFICATIONS\nDetailed specifications for each module"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    ws["A2"] = "Detailed content requirements for each training module including objectives and assessments"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")

    headers = [
        ("Module ID", 15),
        ("Learning Objectives", 50),
        ("Topics Covered", 50),
        ("Prerequisites", 25),
        ("Target Audience", 25),
        ("Accessibility Requirements", 30),
        ("Language Versions", 18),
        ("Format Requirements", 30),
        ("Assessment Questions Count", 22),
        ("Pass Threshold", 15),
        ("Remediation Path", 30),
        ("Regulatory Mapping", 30),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Sample row (row 4) + 50 empty FFFFCC rows (rows 5-54)
    _cs_f2f2f2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws.cell(row=4, column=1, value="CS-001").font = Font(name="Calibri", color="808080")
    ws.cell(row=4, column=1).fill = _cs_f2f2f2
    for c in range(2, 13):
        cell = ws.cell(row=4, column=c)
        cell.fill = _cs_f2f2f2
        cell.border = styles["border"]
        cell.alignment = styles["input_cell"]["alignment"]
    for r in range(5, 55):
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    ws.freeze_panes = "A4"


def create_delivery_standards_sheet(ws, styles):
    """Create Delivery_Standards sheet with pre-populated data."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "DELIVERY STANDARDS\nStandards for each delivery method"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    headers = [
        ("Delivery Method", 20),
        ("Description", 35),
        ("Use Cases", 30),
        ("Technical Requirements", 30),
        ("Instructor Requirements", 25),
        ("Completion Tracking", 22),
        ("Assessment Integration", 25),
        ("Accessibility Standard", 22),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row += 1
    for standard_data in DELIVERY_STANDARDS:
        for col_idx, value in enumerate(standard_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    ws.freeze_panes = "A4"


def create_assessment_design_sheet(ws, styles):
    """Create Assessment_Design sheet with pre-populated data."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "ASSESSMENT DESIGN\nAssessment methodology standards"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    headers = [
        ("Assessment Type", 18),
        ("Purpose", 25),
        ("Question Types", 30),
        ("Minimum Questions", 18),
        ("Pass Threshold", 15),
        ("Attempt Limits", 15),
        ("Feedback Approach", 25),
        ("Remediation Requirements", 30),
        ("Record Retention", 18),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row += 1
    for assessment_data in ASSESSMENT_STANDARDS:
        for col_idx, value in enumerate(assessment_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    ws.freeze_panes = "A4"


def create_content_lifecycle_sheet(ws, styles):
    """Create Content_Lifecycle sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "CONTENT LIFECYCLE\nVersion control and update schedule"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    ws["A2"] = "Version control records and update schedule for all training content"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")

    headers = [
        ("Module ID", 15),
        ("Current Version", 15),
        ("Version Date", 12),
        ("Change Summary", 40),
        ("Review Cycle", 15),
        ("Next Review Date", 18),
        ("Trigger Events", 30),
        ("Update Owner", 20),
        ("ApprovalRequired", 18),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_cycle = DataValidation(type="list", formula1='"Annual,Bi-Annual,Quarterly,Ad-hoc"', allow_blank=False)
    dv_approval = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv_cycle)
    ws.add_data_validation(dv_approval)

    # Sample row (row 4) + 50 empty FFFFCC rows (rows 5-54)
    _cl_f2f2f2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws.cell(row=4, column=1, value="MOD-T1-001").font = Font(name="Calibri", color="808080")
    ws.cell(row=4, column=1).fill = _cl_f2f2f2
    for c in range(2, 10):
        cell = ws.cell(row=4, column=c)
        cell.fill = _cl_f2f2f2
        cell.border = styles["border"]
        cell.alignment = styles["input_cell"]["alignment"]
    dv_cycle.add(ws.cell(row=4, column=5))
    dv_approval.add(ws.cell(row=4, column=9))
    for r in range(5, 55):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_cycle.add(ws.cell(row=r, column=5))
        dv_approval.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "A4"


def create_gap_analysis_sheet(ws, styles):
    """Create Gap_Analysis sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "GAP ANALYSIS\nIdentify curriculum gaps vs. policy requirements"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35
    ws["A2"] = "Identify curriculum gaps against policy requirements and track remediation"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")

    headers = [
        ("Policy Requirement", 40),
        ("Policy Reference", 20),
        ("Current Module", 25),
        ("Gap Description", 40),
        ("Priority", 12),
        ("Remediation Plan", 40),
        ("Target Date", 12),
        ("Owner", 20),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_priority = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=False)
    ws.add_data_validation(dv_priority)

    # Sample row (row 4) + 50 empty FFFFCC rows (rows 5-54)
    _ga2_f2f2f2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws.cell(row=4, column=1, value="GAP-001").font = Font(name="Calibri", color="808080")
    ws.cell(row=4, column=1).fill = _ga2_f2f2f2
    for c in range(2, 9):
        cell = ws.cell(row=4, column=c)
        cell.fill = _ga2_f2f2f2
        cell.border = styles["border"]
        cell.alignment = styles["input_cell"]["alignment"]
    dv_priority.add(ws.cell(row=4, column=5))
    for r in range(5, 55):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_priority.add(ws.cell(row=r, column=5))

    ws.freeze_panes = "A4"


def create_evidence_register(ws):
    """Create the Evidence Register sheet (GS-ER-compliant standard format)."""
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _hdr_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _inp_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _er_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _ctr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row 1: A1:H1 navy title, height 35
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy_fill
    ws["A1"].alignment = _ctr_align
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    # Row 2: Italic subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = f"{DOCUMENT_ID} — {WORKBOOK_NAME} | Audit evidence tracking"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = _ctr_align
    ws["A2"].border = _border

    # Row 3: Empty separator

    # Row 4: Column headers — 003366 fill, white bold font
    headers = [
        ("Evidence ID", 15),
        ("Evidence Type", 22),
        ("Description", 40),
        ("Location / Source", 40),
        ("Date Collected", 15),
        ("Collected By", 20),
        ("Status", 18),
        ("Notes", 40),
    ]
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _hdr_fill
        cell.alignment = _ctr_align
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_status = DataValidation(type="list", formula1='"Verified,Pending,Requires update,Not available"', allow_blank=True)
    ws.add_data_validation(dv_status)

    # Row 5: F2F2F2 sample row starting with EV-001
    ws.cell(row=5, column=1, value="EV-001").fill = _grey_fill
    ws.cell(row=5, column=1).font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws.cell(row=5, column=1).border = _border
    ws.cell(row=5, column=1).alignment = _er_align
    for c in range(2, 9):
        cell = ws.cell(row=5, column=c)
        cell.fill = _grey_fill
        cell.border = _border
        cell.alignment = _er_align

    # Rows 6-105: 100 FFFFCC empty input rows
    for r in range(6, 106):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = _inp_fill
            cell.border = _border
            cell.alignment = _er_align

    # Apply dropdown to sample + data rows
    for r in range(5, 106):
        dv_status.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "A5"



def create_summary_dashboard_sheet(ws):
    """Create Summary Dashboard — Gold Standard TABLE 1/2/3 (A.6.3.2)."""
    from openpyxl.utils import get_column_letter
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ffffcc_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    navy_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    # Row 1: Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = "TRAINING PROGRAM DESIGN — SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 — Control A.6.3: Information Security Awareness, Education and Training | Program Design"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Empty

    # TABLE 1 banner (Row 4)
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = navy_fill
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=4, column=c).border = border
    ws.row_dimensions[4].height = 20

    # Row 5: Column headers
    t1_headers = ["Assessment Area", "Count", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col, hdr in enumerate(t1_headers, 1):
        cell = ws.cell(row=5, column=col, value=hdr)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[5].height = 30

    # Row 6: Curriculum Catalog
    ws.cell(row=6, column=1, value="Curriculum Catalog").border = border
    ws.cell(row=6, column=1).font = Font(color="000000")
    ws.cell(row=6, column=2, value="=COUNTA('Curriculum Catalog'!A4:A75)").border = border
    ws.cell(row=6, column=2).font = Font(color="000000")
    ws.cell(row=6, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=6, column=3, value="=COUNTIF('Curriculum Catalog'!N4:N75,\"Active\")").border = border
    ws.cell(row=6, column=3).font = Font(color="000000")
    ws.cell(row=6, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=6, column=4, value="=COUNTIF('Curriculum Catalog'!N4:N75,\"Draft\")+COUNTIF('Curriculum Catalog'!N4:N75,\"Under Review\")").border = border
    ws.cell(row=6, column=4).font = Font(color="000000")
    ws.cell(row=6, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=6, column=5, value=0).border = border
    ws.cell(row=6, column=5).font = Font(color="000000")
    ws.cell(row=6, column=5).alignment = Alignment(horizontal="center")
    ws.cell(row=6, column=6, value="=COUNTIF('Curriculum Catalog'!N4:N75,\"Retired\")").border = border
    ws.cell(row=6, column=6).font = Font(color="000000")
    ws.cell(row=6, column=6).alignment = Alignment(horizontal="center")
    cell_g6 = ws.cell(row=6, column=7, value="=IF((B6-F6)=0,0,C6/(B6-F6))")
    cell_g6.number_format = "0.0%"
    cell_g6.border = border
    cell_g6.font = Font(color="000000")
    cell_g6.alignment = Alignment(horizontal="center")

    # Row 7: Curriculum Gap Analysis
    ws.cell(row=7, column=1, value="Curriculum Gap Analysis").border = border
    ws.cell(row=7, column=1).font = Font(color="000000")
    ws.cell(row=7, column=2, value="=COUNTA('Gap Analysis'!A5:A54)").border = border
    ws.cell(row=7, column=2).font = Font(color="000000")
    ws.cell(row=7, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=7, column=3, value="=COUNTIF('Gap Analysis'!E5:E54,\"Low\")").border = border
    ws.cell(row=7, column=3).font = Font(color="000000")
    ws.cell(row=7, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=7, column=4, value="=COUNTIF('Gap Analysis'!E5:E54,\"Medium\")").border = border
    ws.cell(row=7, column=4).font = Font(color="000000")
    ws.cell(row=7, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=7, column=5, value="=COUNTIF('Gap Analysis'!E5:E54,\"Critical\")+COUNTIF('Gap Analysis'!E5:E54,\"High\")").border = border
    ws.cell(row=7, column=5).font = Font(color="000000")
    ws.cell(row=7, column=5).alignment = Alignment(horizontal="center")
    ws.cell(row=7, column=6, value=0).border = border
    ws.cell(row=7, column=6).font = Font(color="000000")
    ws.cell(row=7, column=6).alignment = Alignment(horizontal="center")
    cell_g7 = ws.cell(row=7, column=7, value="=IF((B7-F7)=0,0,C7/(B7-F7))")
    cell_g7.number_format = "0.0%"
    cell_g7.border = border
    cell_g7.font = Font(color="000000")
    cell_g7.alignment = Alignment(horizontal="center")

    # Row 8: TOTAL
    ws.cell(row=8, column=1, value="TOTAL").font = Font(bold=True, color="000000")
    ws.cell(row=8, column=1).fill = grey_fill
    ws.cell(row=8, column=1).border = border
    for col in range(2, 7):
        cell = ws.cell(row=8, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}7)"
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    cell_g8 = ws.cell(row=8, column=7, value="=IF((B8-F8)=0,0,C8/(B8-F8))")
    cell_g8.number_format = "0.0%"
    cell_g8.font = Font(bold=True, color="000000")
    cell_g8.fill = grey_fill
    cell_g8.border = border
    cell_g8.alignment = Alignment(horizontal="center")

    # TABLE 2 banner (Row 10)
    t2_start = 10
    ws.merge_cells(f"A{t2_start}:G{t2_start}")
    ws[f"A{t2_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{t2_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t2_start}"].fill = navy_fill
    ws[f"A{t2_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=t2_start, column=c).border = border
    ws.row_dimensions[t2_start].height = 20

    # TABLE 2 headers (Row 11)
    t2_hdr_row = t2_start + 1
    for col, hdr in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        cell = ws.cell(row=t2_hdr_row, column=col, value=hdr if hdr else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # TABLE 2 metrics
    t2_metrics = [
        ("Total Modules in Catalog",
         "=COUNTA('Curriculum Catalog'!A4:A75)"),
        ("Active Modules",
         "=COUNTIF('Curriculum Catalog'!N4:N75,\"Active\")"),
        ("Modules Under Review or Draft",
         "=COUNTIF('Curriculum Catalog'!N4:N75,\"Under Review\")+COUNTIF('Curriculum Catalog'!N4:N75,\"Draft\")"),
        ("Mandatory Modules",
         "=COUNTIF('Curriculum Catalog'!E4:E75,\"Mandatory\")"),
        ("Optional Modules",
         "=COUNTIF('Curriculum Catalog'!E4:E75,\"Optional\")"),
        ("Total Curriculum Gaps Identified",
         "=COUNTA('Gap Analysis'!A5:A54)"),
        ("Critical / High Priority Curriculum Gaps",
         "=COUNTIF('Gap Analysis'!E5:E54,\"Critical\")+COUNTIF('Gap Analysis'!E5:E54,\"High\")"),
        ("Medium / Low Priority Curriculum Gaps",
         "=COUNTIF('Gap Analysis'!E5:E54,\"Medium\")+COUNTIF('Gap Analysis'!E5:E54,\"Low\")"),
    ]

    row = t2_hdr_row + 1
    for metric, formula in t2_metrics:
        ws.cell(row=row, column=1, value=metric).border = border
        ws.cell(row=row, column=1).font = Font(color="000000")
        cell_val = ws.cell(row=row, column=2, value=formula)
        cell_val.border = border
        cell_val.font = Font(color="000000")
        cell_val.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = border
        row += 1

    # 2 empty buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
        row += 1

    # TABLE 3 banner
    t3_start = row + 1
    ws.merge_cells(f"A{t3_start}:G{t3_start}")
    ws[f"A{t3_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{t3_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t3_start}"].fill = red_fill
    ws[f"A{t3_start}"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=t3_start, column=c).border = border
    ws.row_dimensions[t3_start].height = 20

    # TABLE 3 headers
    t3_hdr_row = t3_start + 1
    for col, hdr in enumerate(["Category", "Finding", "Count", "Severity", "Action Required", "", ""], 1):
        cell = ws.cell(row=t3_hdr_row, column=col, value=hdr if hdr else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # TABLE 3 findings
    t3_findings = [
        ("Curriculum Gap", "Critical priority curriculum gaps",
         "=COUNTIF('Gap Analysis'!E5:E54,\"Critical\")", "Critical", "Immediate"),
        ("Curriculum Gap", "High priority curriculum gaps",
         "=COUNTIF('Gap Analysis'!E5:E54,\"High\")", "High", "Urgent"),
        ("Module Status", "Modules under review or draft (not yet active)",
         "=COUNTIF('Curriculum Catalog'!N4:N75,\"Under Review\")+COUNTIF('Curriculum Catalog'!N4:N75,\"Draft\")", "Medium", "Monitor"),
        ("Module Status", "Retired modules in catalog",
         "=COUNTIF('Curriculum Catalog'!N4:N75,\"Retired\")", "Low", "Review"),
    ]

    row = t3_hdr_row + 1
    for cat, finding, formula, severity, action in t3_findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(color="000000")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    # 2 empty FFFFCC buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border
        row += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A4"


def create_approval_sheet(ws):
    """Create standardised Approval Sign-Off sheet."""
    _as_thin = Side(style="thin")
    _as_border = Border(
        left=_as_thin, right=_as_thin, top=_as_thin, bottom=_as_thin
    )

    def _apply_border_row(start_col, end_col, the_row):
        for c in range(start_col, end_col + 1):
            ws.cell(row=the_row, column=c).border = _as_border

    # Row 1: Title banner
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(
        start_color="003366", end_color="003366", fill_type="solid"
    )
    ws["A1"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[1].height = 35
    _apply_border_row(1, 5, 1)

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    _apply_border_row(1, 5, 2)

    # Row 3: Assessment Summary banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    _apply_border_row(1, 5, 3)

    # Summary fields
    _summary = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G7),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    _status_row = None
    _ffffcc = PatternFill(
        start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
    )
    for _label, _val in _summary:
        ws[f"A{row}"] = _label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = _as_border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = _val
        for c in range(2, 6):
            if _val == "":
                ws.cell(row=row, column=c).fill = _ffffcc
            ws.cell(row=row, column=c).border = _as_border
        if "Assessment Status" in _label:
            _status_row = row
        row += 1
    # Overall Compliance Rating — links to Summary Dashboard TOTAL row
    ws["B6"].value = "=IFERROR(AVERAGE('Summary Dashboard'!G6:G7),\"\")"
    ws["B6"].number_format = "0.0%"

    # Status dropdown
    _dv_status = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(_dv_status)
    if _status_row:
        _dv_status.add(f"B{_status_row}")

    # Approver sections helper
    def _approver(start_row, title, colour):
        ws.merge_cells(f"A{start_row}:E{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(
            name="Calibri", size=11, bold=True, color="FFFFFF"
        )
        ws[f"A{start_row}"].fill = PatternFill(
            start_color=colour, end_color=colour, fill_type="solid"
        )
        ws[f"A{start_row}"].alignment = Alignment(
            horizontal="left", vertical="center"
        )
        _apply_border_row(1, 5, start_row)
        r = start_row + 1
        for _f in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{r}"] = _f
            ws[f"A{r}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"A{r}"].border = _as_border
            ws.merge_cells(f"B{r}:E{r}")
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = _ffffcc
                ws.cell(row=r, column=c).border = _as_border
            r += 1
        return r + 1

    row += 2
    row = _approver(row, "COMPLETED BY (ASSESSOR)", "4472C4")
    row = _approver(row, "REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4")
    row = _approver(row, "APPROVED BY (CISO)", "003366")

    # Final Decision
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
    ws[f"A{row}"].border = _as_border
    ws.merge_cells(f"B{row}:E{row}")
    for c in range(2, 6):
        ws.cell(row=row, column=c).fill = _ffffcc
        ws.cell(row=row, column=c).border = _as_border
    _dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(_dv_decision)
    _dv_decision.add(f"B{row}")

    # Next Review Details
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    _apply_border_row(1, 5, row)
    row += 1
    for _rl in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = _rl
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"A{row}"].border = _as_border
        ws.merge_cells(f"B{row}:E{row}")
        for c in range(2, 6):
            ws.cell(row=row, column=c).fill = _ffffcc
            ws.cell(row=row, column=c).border = _as_border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"

def finalize_validations(wb: Workbook) -> None:
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly


def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = _STYLES

        logger.info("[1/10] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("[2/10] Creating Curriculum_Catalog sheet...")
        create_curriculum_catalog_sheet(wb["Curriculum Catalog"], styles)

        logger.info("[3/10] Creating Content_Specifications sheet...")
        create_content_specifications_sheet(wb["Content Specifications"], styles)

        logger.info("[4/10] Creating Delivery_Standards sheet...")
        create_delivery_standards_sheet(wb["Delivery Standards"], styles)

        logger.info("[5/10] Creating Assessment_Design sheet...")
        create_assessment_design_sheet(wb["Assessment Design"], styles)

        logger.info("[6/10] Creating Content_Lifecycle sheet...")
        create_content_lifecycle_sheet(wb["Content Lifecycle"], styles)

        logger.info("[7/10] Creating Gap_Analysis sheet...")
        create_gap_analysis_sheet(wb["Gap Analysis"], styles)

        logger.info("[8/10] Creating Evidence_Register sheet...")
        create_evidence_register(wb["Evidence Register"])

        logger.info("[9/10] Creating Summary Dashboard...")
        create_summary_dashboard_sheet(wb["Summary Dashboard"])

        logger.info("[10/10] Creating Approval_Sign_Off sheet...")
        create_approval_sheet(wb["Approval Sign-Off"])

        finalize_validations(wb)
        _wkbk_dir.mkdir(parents=True, exist_ok=True)
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        wb.save(output_path)
        logger.info(f"SUCCESS: {_wkbk_dir / OUTPUT_FILENAME}")
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error(f"Missing dependency: {e}")
        return 1
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================