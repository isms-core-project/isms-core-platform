#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.3.2 - Training Program Design Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.3: Information Security Awareness, Education and Training
Assessment Domain 2 of 4: Training Program Design

Reference Pattern: Based on ISMS-IMP-A.6.3.2 specification

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates an Excel workbook for documenting Training Program Design
per ISO 27001:2022 Control A.6.3 requirements.

**Purpose:**
Document training curriculum design, content specifications, and delivery
standards to ensure training programs meet policy requirements.

**Generated Workbook Structure:**
1. Instructions - Usage guidance
2. Curriculum_Catalog - Master module listing
3. Content_Specifications - Detailed module specs
4. Delivery_Standards - Delivery method standards
5. Assessment_Design - Assessment methodology
6. Content_Lifecycle - Version control and updates
7. Gap_Analysis - Curriculum gaps vs policy
8. Evidence_Register - Audit evidence
9. Dashboard - Summary metrics
10. Approval_Sign_Off - Formal approval

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.3.2"
WORKBOOK_NAME = "Training Program Design"
CONTROL_ID = "A.6.3"
CONTROL_NAME = "Information Security Awareness, Education and Training"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
    }


# =============================================================================
# PRE-POPULATED DATA
# =============================================================================

CURRICULUM_CATALOG = [
    # Tier 1 - Baseline Awareness
    ("MOD-T1-001", "Information Security Fundamentals", "Tier 1", "Baseline Awareness", "Mandatory", 30, "eLearning", "Quiz", "Onboarding", "POL-A.6.3 S2.3.1"),
    ("MOD-T1-002", "Acceptable Use Policy", "Tier 1", "Baseline Awareness", "Mandatory", 20, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.3.2"),
    ("MOD-T1-003", "Data Classification and Handling", "Tier 1", "Baseline Awareness", "Mandatory", 25, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.3.3"),
    ("MOD-T1-004", "Password Security and Authentication", "Tier 1", "Baseline Awareness", "Mandatory", 20, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.3.4"),
    ("MOD-T1-005", "Phishing and Social Engineering", "Tier 1", "Baseline Awareness", "Mandatory", 30, "eLearning + Simulation", "Quiz + Simulation", "Annual", "POL-A.6.3 S2.3.5"),
    ("MOD-T1-006", "Security Incident Reporting", "Tier 1", "Baseline Awareness", "Mandatory", 15, "eLearning", "Quiz", "Onboarding", "POL-A.6.3 S2.3.6"),
    ("MOD-T1-007", "Physical Security Awareness", "Tier 1", "Baseline Awareness", "Mandatory", 15, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.3.7"),
    ("MOD-T1-008", "Remote Working Security", "Tier 1", "Baseline Awareness", "Mandatory", 20, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.3.8"),
    # Tier 3 - Data Handlers
    ("MOD-T3-001", "Privacy and Data Protection Fundamentals", "Tier 3", "Data Protection", "Mandatory", 45, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.4.1"),
    ("MOD-T3-002", "Data Subject Rights and Obligations", "Tier 3", "Data Protection", "Mandatory", 30, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.4.1"),
    ("MOD-T3-003", "Data Retention and Secure Deletion", "Tier 3", "Data Protection", "Mandatory", 25, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.4.1"),
    # Tier 4 - Technical Staff
    ("MOD-T4-001", "Secure Coding Practices", "Tier 4", "Technical Security", "Mandatory", 60, "eLearning + Workshop", "Quiz + Practical", "Annual", "POL-A.6.3 S2.4.2"),
    ("MOD-T4-002", "OWASP Top 10 Vulnerabilities", "Tier 4", "Technical Security", "Mandatory", 45, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.4.2"),
    ("MOD-T4-003", "Secure Configuration Management", "Tier 4", "Technical Security", "Mandatory", 30, "eLearning", "Quiz", "Annual", "POL-A.6.3 S2.4.2"),
    # Tier 5 - Privileged Users
    ("MOD-T5-001", "Privileged Access Management", "Tier 5", "Privileged Access", "Mandatory", 45, "Instructor-Led", "Quiz + Scenario", "Annual", "POL-A.6.3 S2.4.3"),
    ("MOD-T5-002", "Incident Response for System Admins", "Tier 5", "Privileged Access", "Mandatory", 60, "Workshop", "Practical", "Annual", "POL-A.6.3 S2.4.3"),
    # Tier 6 - Security Roles
    ("MOD-T6-001", "Risk Assessment Methodology", "Tier 6", "Specialized", "Mandatory", 120, "Instructor-Led", "Practical", "Annual", "POL-A.6.3 S2.4.4"),
    ("MOD-T6-002", "Incident Response and Forensics", "Tier 6", "Specialized", "Mandatory", 240, "Workshop", "Tabletop + Practical", "Annual", "POL-A.6.3 S2.4.4"),
    # Tier 7 - Management
    ("MOD-T7-001", "Information Security Governance", "Tier 7", "Governance", "Mandatory", 45, "eLearning + Briefing", "N/A", "Annual", "POL-A.6.3 S2.4.5"),
    ("MOD-T7-002", "Executive Cyber Risk Briefing", "Tier 7", "Governance", "Mandatory", 30, "Briefing", "N/A", "Quarterly", "POL-A.6.3 S2.4.5"),
]

DELIVERY_STANDARDS = [
    ("eLearning", "Self-paced online modules", "Baseline awareness, refresher training", "LMS access, modern browser", "N/A", "Automatic via LMS", "Integrated quiz", "WCAG 2.1 AA"),
    ("Instructor-Led", "Facilitated classroom or virtual", "Complex topics, specialized roles", "Classroom/VC platform", "Certified instructor", "Attendance record", "Post-session assessment", "Accessible venue"),
    ("Workshop", "Hands-on practical exercises", "Technical skills, incident response", "Lab environment", "Technical SME", "Participation record", "Practical assessment", "Adjustable workstations"),
    ("Simulation", "Realistic scenario exercises", "Phishing, social engineering", "Simulation platform", "Campaign design expertise", "Engagement tracking", "Behavioral measurement", "Text alternatives"),
    ("Briefing", "Executive-level sessions", "Leadership updates", "Executive meeting space", "Senior security leadership", "Attendance record", "N/A", "Standard meeting access"),
    ("Microlearning", "Brief focused modules (2-5 min)", "Reinforcement, quick updates", "Mobile-friendly delivery", "N/A", "View tracking", "Optional quick check", "Mobile accessible"),
]

ASSESSMENT_STANDARDS = [
    ("Quiz", "Knowledge verification", "Multiple choice, True/False, Matching", 10, "80%", 3, "Immediate feedback", "Retake training if 3 fails", "5 years"),
    ("Practical", "Skills demonstration", "Hands-on exercises, scenario completion", 3, "Pass/Fail per task", 2, "Instructor feedback", "Additional coaching", "3 years"),
    ("Simulation", "Behavioral assessment", "Phishing response, social engineering", 1, "Did not click + reported", "N/A", "Educational landing page", "Targeted remedial training", "3 years"),
    ("Scenario", "Situational judgment", "Case studies, decision scenarios", 5, "80%", 2, "Explanation of optimal response", "Review with SME", "5 years"),
    ("Tabletop", "Group decision-making", "Incident scenarios, team response", 1, "Participation + contribution", 1, "Facilitator debrief", "Individual follow-up", "3 years"),
]


# =============================================================================
# WORKBOOK CREATION
# =============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Curriculum_Catalog",
        "Content_Specifications",
        "Delivery_Standards",
        "Assessment_Design",
        "Content_Lifecycle",
        "Gap_Analysis",
        "Evidence_Register",
        "Dashboard",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Training Curriculum Framework and Content Management"),
        ("Related Policy", "ISMS-POL-A.6.3, Sections 2.3-2.5"),
        ("Version", "1.0"),
        ("Review Date", ""),
        ("Completed By", ""),
        ("Review Cycle", "Annual + upon significant changes"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Review and customize Curriculum_Catalog with your training modules.",
        "2. Complete Content_Specifications for each module.",
        "3. Review Delivery_Standards for applicable delivery methods.",
        "4. Configure Assessment_Design standards per your requirements.",
        "5. Maintain Content_Lifecycle for version control.",
        "6. Complete Gap_Analysis against policy requirements.",
        "7. Maintain Evidence_Register for audit traceability.",
        "8. Obtain approvals in Approval_Sign_Off.",
    ]

    row += 1
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


def create_curriculum_catalog_sheet(ws, styles):
    """Create Curriculum_Catalog sheet with pre-populated data."""
    ws.merge_cells("A1:N1")
    ws["A1"] = "CURRICULUM CATALOG\nMaster catalog of all training modules"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Module_ID", 15),
        ("Module_Title", 40),
        ("Training_Tier", 12),
        ("Category", 20),
        ("Mandatory_Optional", 18),
        ("Duration_Minutes", 18),
        ("Delivery_Method", 25),
        ("Assessment_Type", 20),
        ("Frequency", 15),
        ("Policy_Reference", 20),
        ("Content_Owner", 20),
        ("Last_Updated", 12),
        ("Next_Review", 12),
        ("Status", 12),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Pre-populate curriculum
    row += 1
    for module_data in CURRICULUM_CATALOG:
        for col_idx, value in enumerate(module_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Add input cells for owner, dates, status
        for col_idx in range(11, 15):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Add empty rows for additional modules
    for r in range(row, row + 30):
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Add dropdowns
    dv_tier = DataValidation(type="list", formula1='"Tier 1,Tier 2,Tier 3,Tier 4,Tier 5,Tier 6,Tier 7"', allow_blank=False)
    dv_mandatory = DataValidation(type="list", formula1='"Mandatory,Optional"', allow_blank=False)
    dv_delivery = DataValidation(type="list", formula1='"eLearning,Instructor-Led,Workshop,Simulation,Briefing,Microlearning,Blended"', allow_blank=False)
    dv_frequency = DataValidation(type="list", formula1='"Onboarding,Annual,Bi-Annual,Quarterly,As Needed"', allow_blank=False)
    dv_status = DataValidation(type="list", formula1='"Active,Draft,Under Review,Retired"', allow_blank=False)

    ws.add_data_validation(dv_tier)
    ws.add_data_validation(dv_mandatory)
    ws.add_data_validation(dv_delivery)
    ws.add_data_validation(dv_frequency)
    ws.add_data_validation(dv_status)

    for r in range(4, row + 30):
        dv_tier.add(ws.cell(row=r, column=3))
        dv_mandatory.add(ws.cell(row=r, column=5))
        dv_delivery.add(ws.cell(row=r, column=7))
        dv_frequency.add(ws.cell(row=r, column=9))
        dv_status.add(ws.cell(row=r, column=14))

    ws.freeze_panes = "A4"


def create_content_specifications_sheet(ws, styles):
    """Create Content_Specifications sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "CONTENT SPECIFICATIONS\nDetailed specifications for each module"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Module_ID", 15),
        ("Learning_Objectives", 50),
        ("Topics_Covered", 50),
        ("Prerequisites", 25),
        ("Target_Audience", 25),
        ("Accessibility_Requirements", 30),
        ("Language_Versions", 18),
        ("Format_Requirements", 30),
        ("Assessment_Questions_Count", 22),
        ("Pass_Threshold", 15),
        ("Remediation_Path", 30),
        ("Regulatory_Mapping", 30),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for r in range(4, 54):
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

    ws.freeze_panes = "A4"


def create_delivery_standards_sheet(ws, styles):
    """Create Delivery_Standards sheet with pre-populated data."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "DELIVERY STANDARDS\nStandards for each delivery method"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Delivery_Method", 20),
        ("Description", 35),
        ("Use_Cases", 30),
        ("Technical_Requirements", 30),
        ("Instructor_Requirements", 25),
        ("Completion_Tracking", 22),
        ("Assessment_Integration", 25),
        ("Accessibility_Standard", 22),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row += 1
    for standard_data in DELIVERY_STANDARDS:
        for col_idx, value in enumerate(standard_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    ws.freeze_panes = "A4"


def create_assessment_design_sheet(ws, styles):
    """Create Assessment_Design sheet with pre-populated data."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "ASSESSMENT DESIGN\nAssessment methodology standards"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Assessment_Type", 18),
        ("Purpose", 25),
        ("Question_Types", 30),
        ("Minimum_Questions", 18),
        ("Pass_Threshold", 15),
        ("Attempt_Limits", 15),
        ("Feedback_Approach", 25),
        ("Remediation_Requirements", 30),
        ("Record_Retention", 18),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row += 1
    for assessment_data in ASSESSMENT_STANDARDS:
        for col_idx, value in enumerate(assessment_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    ws.freeze_panes = "A4"


def create_content_lifecycle_sheet(ws, styles):
    """Create Content_Lifecycle sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "CONTENT LIFECYCLE\nVersion control and update schedule"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Module_ID", 15),
        ("Current_Version", 15),
        ("Version_Date", 12),
        ("Change_Summary", 40),
        ("Review_Cycle", 15),
        ("Next_Review_Date", 18),
        ("Trigger_Events", 30),
        ("Update_Owner", 20),
        ("Approval_Required", 18),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_cycle = DataValidation(type="list", formula1='"Annual,Bi-Annual,Quarterly,Ad-hoc"', allow_blank=False)
    dv_approval = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv_cycle)
    ws.add_data_validation(dv_approval)

    for r in range(4, 54):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_cycle.add(ws.cell(row=r, column=5))
        dv_approval.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "A4"


def create_gap_analysis_sheet(ws, styles):
    """Create Gap_Analysis sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "GAP ANALYSIS\nIdentify curriculum gaps vs. policy requirements"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Policy_Requirement", 40),
        ("Policy_Reference", 20),
        ("Current_Module", 25),
        ("Gap_Description", 40),
        ("Priority", 12),
        ("Remediation_Plan", 40),
        ("Target_Date", 12),
        ("Owner", 20),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_priority = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=False)
    ws.add_data_validation(dv_priority)

    for r in range(4, 54):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_priority.add(ws.cell(row=r, column=5))

    ws.freeze_panes = "A4"


def create_evidence_register_sheet(ws, styles):
    """Create Evidence_Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 40),
        ("Location", 40),
        ("Date_Collected", 15),
        ("Collected_By", 20),
        ("Status", 18),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_status = DataValidation(type="list", formula1='"Verified,Pending,Requires update,Not available"', allow_blank=False)
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_status.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "A4"


def create_dashboard_sheet(ws, styles):
    """Create Dashboard sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "TRAINING PROGRAM DESIGN - DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Curriculum Summary"
    ws["A3"].font = Font(bold=True, size=12)

    metrics = [
        ("Total Modules", '=COUNTA(Curriculum_Catalog!A4:A53)'),
        ("Active Modules", '=COUNTIF(Curriculum_Catalog!N4:N53,"Active")'),
        ("Modules Under Review", '=COUNTIF(Curriculum_Catalog!N4:N53,"Under Review")'),
        ("Mandatory Modules", '=COUNTIF(Curriculum_Catalog!E4:E53,"Mandatory")'),
        ("Optional Modules", '=COUNTIF(Curriculum_Catalog!E4:E53,"Optional")'),
    ]

    row = 4
    for label, formula in metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF", size=14)
        row += 1

    row += 2
    ws[f"A{row}"] = "Modules by Tier"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    headers = ["Tier", "Module Count"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for tier_num in range(1, 8):
        ws.cell(row=row, column=1, value=f"Tier {tier_num}").border = styles["border"]
        ws.cell(row=row, column=2, value=f'=COUNTIF(Curriculum_Catalog!C4:C53,"Tier {tier_num}")').border = styles["border"]
        row += 1

    row += 2
    ws[f"A{row}"] = "Gap Analysis Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    gap_metrics = [
        ("Total Gaps Identified", '=COUNTA(Gap_Analysis!A4:A53)'),
        ("Critical Gaps", '=COUNTIF(Gap_Analysis!E4:E53,"Critical")'),
        ("High Priority Gaps", '=COUNTIF(Gap_Analysis!E4:E53,"High")'),
    ]

    for label, formula in gap_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def create_approval_signoff_sheet(ws, styles):
    """Create Approval_Sign_Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Total Modules Documented", "=Dashboard!B4"),
        ("Assessment Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    dv_status = DataValidation(type="list", formula1='"Draft,Final,Requires revision"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row-1}"])

    # Approval sections
    sections = [
        ("REVIEWED BY (Training Program Manager)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    for section_title, color in sections:
        row += 2
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = section_title
        ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

        fields = ["Name", "Date", "Signature"] if "REVIEWED" in section_title else ["Name", "Date", "Approval Decision", "Signature"]
        row += 1
        for field in fields:
            ws[f"A{row}"] = field + ":"
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
            row += 1

    dv_decision = DataValidation(type="list", formula1='"Approved,Approved with conditions,Rejected"', allow_blank=False)
    ws.add_data_validation(dv_decision)
    dv_decision.add(ws[f"B{row-2}"])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/10] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/10] Creating Curriculum_Catalog sheet...")
        create_curriculum_catalog_sheet(wb["Curriculum_Catalog"], styles)

        logger.info("[3/10] Creating Content_Specifications sheet...")
        create_content_specifications_sheet(wb["Content_Specifications"], styles)

        logger.info("[4/10] Creating Delivery_Standards sheet...")
        create_delivery_standards_sheet(wb["Delivery_Standards"], styles)

        logger.info("[5/10] Creating Assessment_Design sheet...")
        create_assessment_design_sheet(wb["Assessment_Design"], styles)

        logger.info("[6/10] Creating Content_Lifecycle sheet...")
        create_content_lifecycle_sheet(wb["Content_Lifecycle"], styles)

        logger.info("[7/10] Creating Gap_Analysis sheet...")
        create_gap_analysis_sheet(wb["Gap_Analysis"], styles)

        logger.info("[8/10] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[9/10] Creating Dashboard sheet...")
        create_dashboard_sheet(wb["Dashboard"], styles)

        logger.info("[10/10] Creating Approval_Sign_Off sheet...")
        create_approval_signoff_sheet(wb["Approval_Sign_Off"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation following ISMS-IMP-A.6.3.2 specification
# =============================================================================
