#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.3.1 - Training Needs Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.3: Information Security Awareness, Education and Training
Assessment Domain 1 of 4: Training Needs Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific role definitions, training requirements, and
competency assessment criteria.

Key customization areas:
1. Role definitions and training tier assignments
2. Training topic requirements per tier
3. Competency scoring criteria
4. Integration with HRIS and LMS systems
5. Regulatory training requirements specific to your jurisdiction

Reference Pattern: Based on ISMS-IMP-A.6.3.1 specification

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates an Excel workbook for conducting Training Needs Assessment
per ISO 27001:2022 Control A.6.3 requirements.

**Purpose:**
Enables systematic identification of training requirements based on job roles,
responsibilities, and information security risk exposure.

**Assessment Scope:**
- Role inventory and classification
- Training tier assignment (Tiers 1-7)
- Training requirements mapping
- Gap analysis between required and current training
- Competency assessment tracking

**Generated Workbook Structure:**
1. Instructions - Assessment guidance
2. Role_Inventory - Complete role listing
3. Tier_Classification - Tier assignment criteria and scoring
4. Training_Requirements - Requirements by tier
5. Gap_Analysis - Current vs required training gaps
6. Evidence_Register - Audit evidence tracking
7. Dashboard - Summary metrics
8. Approval_Sign_Off - Stakeholder approval

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    pip3 install openpyxl

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.6.3
Assessment Domain:    1 of 4 (Training Needs Assessment)
Framework Version:    1.0
Script Version:       1.0
Author:               ISMS Core Contributors
Created:              2026
Last Modified:        [Date to be set]

Related Documents:
    - ISMS-POL-A.6.3: Information Security Awareness and Training Policy
    - ISMS-IMP-A.6.3.2: Training Program Design
    - ISMS-IMP-A.6.3.3: Training Delivery and Tracking
    - ISMS-IMP-A.6.3.4: Training Compliance Dashboard

================================================================================
"""

# =============================================================================
# IMPORTS - Standard Library
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# IMPORTS - Third Party
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.3.1"
WORKBOOK_NAME = "Training Needs Assessment"
CONTROL_ID = "A.6.3"
CONTROL_NAME = "Information Security Awareness, Education and Training"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # Swiss format
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# =============================================================================
# PRE-POPULATED DATA
# =============================================================================

# Training Tiers per POL-A.6.3
TRAINING_TIERS = [
    ("Tier 1", "Baseline Awareness", "All Personnel", "Foundational security awareness required for everyone"),
    ("Tier 2", "Contractor/Third-Party", "External Personnel", "Additional controls for non-employee access"),
    ("Tier 3", "Data Handlers", "Personnel handling sensitive data", "Enhanced training for data protection responsibilities"),
    ("Tier 4", "Technical Staff", "IT/Development teams", "Technical security skills and secure development"),
    ("Tier 5", "Privileged Users", "System/Database administrators", "Advanced security for privileged access holders"),
    ("Tier 6", "Security Roles", "Security team members", "Specialized security operations training"),
    ("Tier 7", "Management", "Executives and managers", "Governance, risk oversight, and leadership responsibilities"),
]

# Tier Assignment Scoring Criteria
TIER_SCORING = [
    ("Data Access Level", "None=0, Public=1, Internal=2, Confidential=3, Restricted=4"),
    ("System Privileges", "None=0, User=1, Power User=2, Admin=3, Super Admin=4"),
    ("External Exposure", "None=0, Internal Only=1, Partner Access=2, Customer-Facing=3, Public-Facing=4"),
    ("Regulatory Scope", "None=0, General=1, Industry-Specific=2, Multiple Regulations=3, Critical Infrastructure=4"),
    ("Security Role", "None=0, Awareness Only=1, Incident Reporter=2, Security Champion=3, Security Team=4"),
]

# Training Requirements by Tier
TRAINING_REQUIREMENTS = [
    # Tier 1 - Baseline
    ("Tier 1", "MOD-T1-001", "Information Security Fundamentals", "Mandatory", "Onboarding", 30),
    ("Tier 1", "MOD-T1-002", "Acceptable Use Policy", "Mandatory", "Annual", 20),
    ("Tier 1", "MOD-T1-003", "Data Classification and Handling", "Mandatory", "Annual", 25),
    ("Tier 1", "MOD-T1-004", "Password Security and Authentication", "Mandatory", "Annual", 20),
    ("Tier 1", "MOD-T1-005", "Phishing and Social Engineering", "Mandatory", "Annual", 30),
    ("Tier 1", "MOD-T1-006", "Security Incident Reporting", "Mandatory", "Onboarding", 15),
    ("Tier 1", "MOD-T1-007", "Physical Security Awareness", "Mandatory", "Annual", 15),
    ("Tier 1", "MOD-T1-008", "Remote Working Security", "Mandatory", "Annual", 20),
    # Tier 3 - Data Handlers
    ("Tier 3", "MOD-T3-001", "Privacy and Data Protection Fundamentals", "Mandatory", "Annual", 45),
    ("Tier 3", "MOD-T3-002", "Data Subject Rights and Obligations", "Mandatory", "Annual", 30),
    ("Tier 3", "MOD-T3-003", "Data Retention and Secure Deletion", "Mandatory", "Annual", 25),
    # Tier 4 - Technical Staff
    ("Tier 4", "MOD-T4-001", "Secure Coding Practices", "Mandatory", "Annual", 60),
    ("Tier 4", "MOD-T4-002", "OWASP Top 10 Vulnerabilities", "Mandatory", "Annual", 45),
    ("Tier 4", "MOD-T4-003", "Secure Configuration Management", "Mandatory", "Annual", 30),
    # Tier 5 - Privileged Users
    ("Tier 5", "MOD-T5-001", "Privileged Access Management", "Mandatory", "Annual", 45),
    ("Tier 5", "MOD-T5-002", "Incident Response for System Admins", "Mandatory", "Annual", 60),
    # Tier 6 - Security Roles
    ("Tier 6", "MOD-T6-001", "Risk Assessment Methodology", "Mandatory", "Annual", 120),
    ("Tier 6", "MOD-T6-002", "Incident Response and Forensics", "Mandatory", "Annual", 240),
    # Tier 7 - Management
    ("Tier 7", "MOD-T7-001", "Information Security Governance", "Mandatory", "Annual", 45),
    ("Tier 7", "MOD-T7-002", "Executive Cyber Risk Briefing", "Mandatory", "Quarterly", 30),
]


# =============================================================================
# WORKBOOK CREATION
# =============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Role_Inventory",
        "Tier_Classification",
        "Training_Requirements",
        "Gap_Analysis",
        "Evidence_Register",
        "Dashboard",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# SHEET: INSTRUCTIONS
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create the Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Training Needs Assessment"),
        ("Related Policy", "ISMS-POL-A.6.3, Section 2.1-2.2"),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organization", ""),
        ("Review Cycle", "Annual + upon significant organizational changes"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Complete Role_Inventory with all job roles in your organization.",
        "2. Use Tier_Classification to assign training tiers based on scoring criteria.",
        "3. Review Training_Requirements for applicable modules per tier.",
        "4. Complete Gap_Analysis to identify training gaps per role.",
        "5. Maintain Evidence_Register for audit traceability.",
        "6. Review Dashboard for summary metrics.",
        "7. Obtain approvals in Approval_Sign_Off sheet.",
    ]

    row += 1
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    row += 1
    ws[f"A{row}"] = "TRAINING TIER OVERVIEW"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    tier_headers = ["Tier", "Name", "Audience", "Description"]
    for col_idx, header in enumerate(tier_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for tier_data in TRAINING_TIERS:
        for col_idx, value in enumerate(tier_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 50


# =============================================================================
# SHEET: ROLE_INVENTORY
# =============================================================================

def create_role_inventory_sheet(ws, styles):
    """Create the Role_Inventory sheet."""
    ws.merge_cells("A1:L1")
    ws["A1"] = "ROLE INVENTORY\nComplete listing of all job roles requiring training"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Role_ID", 12),
        ("Role_Title", 30),
        ("Department", 20),
        ("Employee_Count", 15),
        ("Data_Access_Level", 18),
        ("System_Privileges", 18),
        ("External_Exposure", 18),
        ("Regulatory_Scope", 18),
        ("Security_Role", 18),
        ("Total_Score", 12),
        ("Assigned_Tier", 15),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validation dropdowns
    dv_data_access = DataValidation(
        type="list",
        formula1='"None (0),Public (1),Internal (2),Confidential (3),Restricted (4)"',
        allow_blank=False
    )
    dv_privileges = DataValidation(
        type="list",
        formula1='"None (0),User (1),Power User (2),Admin (3),Super Admin (4)"',
        allow_blank=False
    )
    dv_exposure = DataValidation(
        type="list",
        formula1='"None (0),Internal Only (1),Partner Access (2),Customer-Facing (3),Public-Facing (4)"',
        allow_blank=False
    )
    dv_regulatory = DataValidation(
        type="list",
        formula1='"None (0),General (1),Industry-Specific (2),Multiple Regulations (3),Critical Infrastructure (4)"',
        allow_blank=False
    )
    dv_security = DataValidation(
        type="list",
        formula1='"None (0),Awareness Only (1),Incident Reporter (2),Security Champion (3),Security Team (4)"',
        allow_blank=False
    )

    ws.add_data_validation(dv_data_access)
    ws.add_data_validation(dv_privileges)
    ws.add_data_validation(dv_exposure)
    ws.add_data_validation(dv_regulatory)
    ws.add_data_validation(dv_security)

    # Data rows
    for r in range(4, 104):
        ws.cell(row=r, column=1, value=f"ROLE-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Apply dropdowns
        dv_data_access.add(ws.cell(row=r, column=5))
        dv_privileges.add(ws.cell(row=r, column=6))
        dv_exposure.add(ws.cell(row=r, column=7))
        dv_regulatory.add(ws.cell(row=r, column=8))
        dv_security.add(ws.cell(row=r, column=9))

        # Total Score formula
        ws.cell(row=r, column=10, value=f'=IF(B{r}="","",VALUE(LEFT(E{r},1))+VALUE(LEFT(F{r},1))+VALUE(LEFT(G{r},1))+VALUE(LEFT(H{r},1))+VALUE(LEFT(I{r},1)))')
        ws.cell(row=r, column=10).font = Font(bold=True, color="0000FF")

        # Assigned Tier formula
        ws.cell(row=r, column=11, value=f'=IF(J{r}="","",IF(I{r}="Security Team (4)","Tier 6",IF(OR(F{r}="Admin (3)",F{r}="Super Admin (4)"),"Tier 5",IF(F{r}="Power User (2)","Tier 4",IF(E{r}="Confidential (3)","Tier 3",IF(J{r}>=8,"Tier 7","Tier 1"))))))')

    ws.freeze_panes = "A4"


# =============================================================================
# SHEET: TIER_CLASSIFICATION
# =============================================================================

def create_tier_classification_sheet(ws, styles):
    """Create the Tier_Classification sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "TIER CLASSIFICATION CRITERIA\nScoring matrix for training tier assignment"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Scoring Criteria"
    ws["A3"].font = Font(bold=True, size=12)

    headers = ["Criterion", "Scoring Scale"]
    row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for criterion, scale in TIER_SCORING:
        ws.cell(row=row, column=1, value=criterion).border = styles["border"]
        ws.cell(row=row, column=2, value=scale).border = styles["border"]
        row += 1

    row += 2
    ws[f"A{row}"] = "Tier Assignment Logic"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    tier_logic = [
        ("Tier 7 - Management", "Total Score >= 8 AND in management role"),
        ("Tier 6 - Security Roles", "Security Role = Security Team (4)"),
        ("Tier 5 - Privileged Users", "System Privileges = Admin (3) or Super Admin (4)"),
        ("Tier 4 - Technical Staff", "System Privileges = Power User (2)"),
        ("Tier 3 - Data Handlers", "Data Access Level = Confidential (3) or higher"),
        ("Tier 2 - Contractor", "Employment Type = Contractor/Third-Party"),
        ("Tier 1 - Baseline", "All other personnel (default)"),
    ]

    row += 1
    headers = ["Tier", "Assignment Criteria"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for tier, criteria in tier_logic:
        ws.cell(row=row, column=1, value=tier).border = styles["border"]
        ws.cell(row=row, column=2, value=criteria).border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


# =============================================================================
# SHEET: TRAINING_REQUIREMENTS
# =============================================================================

def create_training_requirements_sheet(ws, styles):
    """Create the Training_Requirements sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "TRAINING REQUIREMENTS BY TIER\nRequired modules per training tier"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Tier", 12),
        ("Module_ID", 15),
        ("Module_Title", 40),
        ("Mandatory_Optional", 18),
        ("Frequency", 15),
        ("Duration_Minutes", 18),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row += 1
    current_tier = ""
    for tier, module_id, title, mandatory, frequency, duration in TRAINING_REQUIREMENTS:
        # Add tier separator
        if tier != current_tier:
            if current_tier != "":
                row += 1  # Empty row between tiers
            current_tier = tier

        ws.cell(row=row, column=1, value=tier).border = styles["border"]
        ws.cell(row=row, column=2, value=module_id).border = styles["border"]
        ws.cell(row=row, column=3, value=title).border = styles["border"]
        ws.cell(row=row, column=4, value=mandatory).border = styles["border"]
        ws.cell(row=row, column=5, value=frequency).border = styles["border"]
        ws.cell(row=row, column=6, value=duration).border = styles["border"]

        # Color code by tier
        tier_colors = {
            "Tier 1": "E2EFDA",  # Light green
            "Tier 3": "DDEBF7",  # Light blue
            "Tier 4": "FFF2CC",  # Light yellow
            "Tier 5": "FCE4D6",  # Light orange
            "Tier 6": "E4DFEC",  # Light purple
            "Tier 7": "F2F2F2",  # Light gray
        }
        if tier in tier_colors:
            fill = PatternFill(start_color=tier_colors[tier], end_color=tier_colors[tier], fill_type="solid")
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = fill

        row += 1

    ws.freeze_panes = "A4"


# =============================================================================
# SHEET: GAP_ANALYSIS
# =============================================================================

def create_gap_analysis_sheet(ws, styles):
    """Create the Gap_Analysis sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "TRAINING GAP ANALYSIS\nIdentify gaps between required and current training status"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Role_ID", 12),
        ("Role_Title", 30),
        ("Assigned_Tier", 15),
        ("Required_Modules", 20),
        ("Completed_Modules", 20),
        ("Gap_Count", 12),
        ("Gap_Percentage", 15),
        ("Priority", 12),
        ("Remediation_Plan", 40),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Priority dropdown
    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_priority)

    # Data rows
    for r in range(4, 104):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Gap Count formula
        ws.cell(row=r, column=6, value=f'=IF(D{r}="","",D{r}-E{r})')
        ws.cell(row=r, column=6).font = Font(bold=True)

        # Gap Percentage formula
        ws.cell(row=r, column=7, value=f'=IF(D{r}="","",IF(D{r}=0,"0%",ROUND((D{r}-E{r})/D{r}*100,1)&"%"))')
        ws.cell(row=r, column=7).font = Font(bold=True, color="0000FF")

        dv_priority.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A4"


# =============================================================================
# SHEET: EVIDENCE_REGISTER
# =============================================================================

def create_evidence_register_sheet(ws, styles):
    """Create the Evidence_Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 40),
        ("Location", 40),
        ("Date_Collected", 15),
        ("Collected_By", 20),
        ("Status", 18),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Dropdowns
    dv_type = DataValidation(
        type="list",
        formula1='"Role documentation,Job descriptions,Org chart,Training records,HRIS export,LMS export,Interview notes,Other"',
        allow_blank=False
    )
    dv_status = DataValidation(
        type="list",
        formula1='"Verified,Pending,Requires update,Not available"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        dv_type.add(ws.cell(row=r, column=2))
        dv_status.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "A4"


# =============================================================================
# SHEET: DASHBOARD
# =============================================================================

def create_dashboard_sheet(ws, styles):
    """Create the Dashboard sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "TRAINING NEEDS ASSESSMENT - DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    # Summary metrics
    ws["A3"] = "Summary Metrics"
    ws["A3"].font = Font(bold=True, size=12)

    metrics = [
        ("Total Roles Assessed", '=COUNTA(Role_Inventory!B4:B103)'),
        ("Total Employees Covered", '=SUM(Role_Inventory!D4:D103)'),
        ("Roles with Training Gaps", '=COUNTIF(Gap_Analysis!F4:F103,">0")'),
        ("Critical Priority Gaps", '=COUNTIF(Gap_Analysis!H4:H103,"Critical")'),
        ("High Priority Gaps", '=COUNTIF(Gap_Analysis!H4:H103,"High")'),
    ]

    row = 4
    for label, formula in metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF", size=14)
        row += 1

    # Tier Distribution
    row += 2
    ws[f"A{row}"] = "Tier Distribution"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    headers = ["Tier", "Role Count", "Employee Count"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for tier, _, _, _ in TRAINING_TIERS:
        ws.cell(row=row, column=1, value=tier).border = styles["border"]
        ws.cell(row=row, column=2, value=f'=COUNTIF(Role_Inventory!K4:K103,"{tier}")').border = styles["border"]
        ws.cell(row=row, column=3, value=f'=SUMIF(Role_Inventory!K4:K103,"{tier}",Role_Inventory!D4:D103)').border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20


# =============================================================================
# SHEET: APPROVAL_SIGN_OFF
# =============================================================================

def create_approval_signoff_sheet(ws, styles):
    """Create the Approval_Sign_Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Assessment Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Assessment Document", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period", ""),
        ("Total Roles Assessed", "=Dashboard!B4"),
        ("Assessment Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    # Status dropdown
    dv_status = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires revision"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row-1}"])

    # Completed By section
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT COMPLETED BY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    completion_fields = ["Name", "Role/Title", "Department", "Email", "Date"]
    row += 1
    for field in completion_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Reviewed By section
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "REVIEWED BY (HR / Training Manager)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    review_fields = ["Name", "Date", "Signature"]
    row += 1
    for field in review_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Approved By section
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "APPROVED BY (CISO)"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

    approval_fields = ["Name", "Date", "Approval Decision", "Signature"]
    row += 1
    for field in approval_fields:
        ws[f"A{row}"] = field + ":"
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    # Approval Decision dropdown
    dv_decision = DataValidation(
        type="list",
        formula1='"Approved,Approved with conditions,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_decision)
    dv_decision.add(ws[f"B{row-2}"])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/8] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/8] Creating Role_Inventory sheet...")
        create_role_inventory_sheet(wb["Role_Inventory"], styles)

        logger.info("[3/8] Creating Tier_Classification sheet...")
        create_tier_classification_sheet(wb["Tier_Classification"], styles)

        logger.info("[4/8] Creating Training_Requirements sheet...")
        create_training_requirements_sheet(wb["Training_Requirements"], styles)

        logger.info("[5/8] Creating Gap_Analysis sheet...")
        create_gap_analysis_sheet(wb["Gap_Analysis"], styles)

        logger.info("[6/8] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[7/8] Creating Dashboard sheet...")
        create_dashboard_sheet(wb["Dashboard"], styles)

        logger.info("[8/8] Creating Approval_Sign_Off sheet...")
        create_approval_signoff_sheet(wb["Approval_Sign_Off"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        logger.error("Install with: pip install openpyxl")
        return 1
    except PermissionError as e:
        logger.error("Permission denied: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation following ISMS-IMP-A.6.3.1 specification
# =============================================================================
