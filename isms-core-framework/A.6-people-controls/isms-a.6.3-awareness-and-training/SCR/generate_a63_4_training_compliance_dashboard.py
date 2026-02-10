#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.6.3.4 - Training Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.6.3: Information Security Awareness, Education and Training
Assessment Domain 4 of 4: Training Compliance Dashboard

Reference Pattern: Based on ISMS-IMP-A.6.3.4 specification

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates an Excel workbook for consolidating training program data
to provide executive-level compliance visibility, effectiveness measurement,
and audit-ready reporting.

**Purpose:**
Provide executive-level compliance visibility, effectiveness measurement,
and audit-ready reporting per ISMS-POL-A.6.3 requirements.

**Generated Workbook Structure:**
1. Instructions - Dashboard usage guidance
2. Executive_Summary - High-level compliance status
3. Compliance_Metrics - Detailed compliance breakdowns
4. Effectiveness_Metrics - Training impact measurements
5. Trend_Analysis - Historical performance tracking
6. Risk_Heatmap - Visual risk identification
7. Audit_Evidence - Evidence checklist for audits
8. Management_Review_Input - Prepared input for management reviews
9. KPI_Definitions - Reference for all metrics
10. Data_Sources - Document data inputs and refresh procedures
11. Evidence_Register - Supporting documentation
12. Dashboard - Visual summary with charts
13. Approval_Sign_Off - Formal review and attestation

================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.6.3.4"
WORKBOOK_NAME = "Training Compliance Dashboard"
CONTROL_ID = "A.6.3"
CONTROL_NAME = "Information Security Awareness, Education and Training"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "green": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "yellow": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "red": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "orange": {"fill": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")},
    }


# =============================================================================
# PRE-POPULATED DATA
# =============================================================================

EFFECTIVENESS_METRICS = [
    ("Behaviour", "Phishing Click Rate", "≤5%", "IMP-A.6.3.3 Simulation_Results"),
    ("Behaviour", "Phishing Report Rate", "≥70%", "IMP-A.6.3.3 Simulation_Results"),
    ("Behaviour", "Credential Submission Rate", "≤1%", "IMP-A.6.3.3 Simulation_Results"),
    ("Knowledge", "First-Attempt Pass Rate", "≥85%", "IMP-A.6.3.3 Assessment_Results"),
    ("Knowledge", "Average Assessment Score", "≥80%", "IMP-A.6.3.3 Assessment_Results"),
    ("Knowledge", "Remediation Success Rate", "≥95%", "IMP-A.6.3.3 Remediation_Tracking"),
    ("Compliance", "On-Time Completion Rate", "≥90%", "IMP-A.6.3.3 Completion_Tracking"),
    ("Compliance", "Annual Training Completion", "100%", "IMP-A.6.3.3 Completion_Tracking"),
    ("Risk", "Training-Related Incidents", "Decreasing", "Incident Management System"),
    ("Risk", "Policy Violations (Training Gap)", "Decreasing", "Policy Violation Log"),
]

AUDIT_EVIDENCE = [
    # Stage 1 - Documentation
    ("Stage 1", "Training Policy (ISMS-POL-A.6.3)", "ISO 27001 A.6.3", "CISO"),
    ("Stage 1", "Training Needs Assessment (IMP-A.6.3.1)", "ISO 27001 A.6.3", "ISO/Training Manager"),
    ("Stage 1", "Training Curriculum Catalog (IMP-A.6.3.2)", "ISO 27001 A.6.3", "Training Manager"),
    ("Stage 1", "Role-to-Training Matrix", "ISO 27001 A.6.3", "HR/Security"),
    ("Stage 1", "Training Delivery Standards", "ISO 27001 A.6.3", "Training Manager"),
    # Stage 2 - Operational
    ("Stage 2", "Training Completion Records", "ISO 27001 Cl.7.2(d)", "HR/LMS Admin"),
    ("Stage 2", "Assessment Results and Scores", "ISO 27001 Cl.7.2(d)", "LMS Admin"),
    ("Stage 2", "Phishing Simulation Results", "ISO 27001 A.6.3", "Security Team"),
    ("Stage 2", "Remediation Tracking Records", "ISO 27001 A.6.3", "HR/Security"),
    ("Stage 2", "Compliance Reports (Monthly)", "ISO 27001 Cl.9.1", "ISO"),
    ("Stage 2", "Effectiveness Metrics (Quarterly)", "ISO 27001 Cl.9.1", "ISO"),
    ("Stage 2", "Management Review Minutes", "ISO 27001 Cl.9.3", "CISO"),
    ("Stage 2", "Training Program Review Records", "ISO 27001 Cl.10.1", "Training Manager"),
]

KPI_DEFINITIONS = [
    ("KPI-001", "Training Completion Rate", "Percentage of required training completed", "(Completed / Required) × 100", "≥95%", "Monthly"),
    ("KPI-002", "On-Time Completion Rate", "Percentage completed before due date", "(On-Time / Completed) × 100", "≥90%", "Monthly"),
    ("KPI-003", "Assessment Pass Rate", "First-attempt pass rate", "(Passed First Attempt / Total) × 100", "≥85%", "Monthly"),
    ("KPI-004", "Phishing Click Rate", "Percentage clicking simulation links", "(Clicked / Sent) × 100", "≤5%", "Per Campaign"),
    ("KPI-005", "Phishing Report Rate", "Percentage reporting suspicious emails", "(Reported / Sent) × 100", "≥70%", "Per Campaign"),
    ("KPI-006", "Remediation Success Rate", "Percentage passing remediation training", "(Remediation Passed / Assigned) × 100", "≥95%", "Monthly"),
    ("KPI-007", "Training Currency", "Percentage with current/non-expired training", "(Current / Total Required) × 100", "≥95%", "Monthly"),
    ("KPI-008", "Average Assessment Score", "Mean score across all assessments", "AVG(Assessment_Score)", "≥80%", "Quarterly"),
]


# =============================================================================
# WORKBOOK CREATION
# =============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheets = [
        "Instructions",
        "Executive_Summary",
        "Compliance_Metrics",
        "Effectiveness_Metrics",
        "Trend_Analysis",
        "Risk_Heatmap",
        "Audit_Evidence",
        "Management_Review_Input",
        "KPI_Definitions",
        "Data_Sources",
        "Evidence_Register",
        "Dashboard",
        "Approval_Sign_Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(ws, styles):
    """Create Instructions sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Document Information"
    ws["A3"].font = Font(bold=True, size=12)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Assessment Area", "Training Program Effectiveness and Compliance Reporting"),
        ("Related Policy", "ISMS-POL-A.6.3, Sections 3-4"),
        ("Version", "1.0"),
        ("Reporting Period", ""),
        ("Prepared By", ""),
        ("Review Cycle", "Monthly reporting + Quarterly management review"),
    ]

    row = 4
    for label, value in doc_info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
        row += 1

    row += 1
    ws[f"A{row}"] = "HOW TO USE THIS WORKBOOK"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    instructions = [
        "1. Populate Compliance_Metrics with data from IMP-A.6.3.3 workbook.",
        "2. Update Effectiveness_Metrics quarterly.",
        "3. Maintain Trend_Analysis with monthly data points.",
        "4. Update Risk_Heatmap based on current compliance status.",
        "5. Verify Audit_Evidence checklist completeness.",
        "6. Prepare Management_Review_Input before management meetings.",
        "7. Review KPI_Definitions for metric understanding.",
        "8. Document data refresh schedules in Data_Sources.",
        "9. Present Executive_Summary to leadership.",
        "10. Complete Approval_Sign_Off for attestation.",
    ]

    row += 1
    for line in instructions:
        ws[f"A{row}"] = line
        row += 1

    row += 1
    ws[f"A{row}"] = "TRAFFIC LIGHT LEGEND"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    legend = [
        ("Green", "≥95% compliance, <5% click rate", "C6EFCE"),
        ("Yellow", "85-94% compliance, 5-15% click rate", "FFEB9C"),
        ("Red", "<85% compliance, >15% click rate", "FFC7CE"),
    ]

    row += 1
    for status, description, color in legend:
        ws[f"A{row}"] = status
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{row}"].border = styles["border"]
        ws[f"B{row}"] = description
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


def create_executive_summary_sheet(ws, styles):
    """Create Executive_Summary sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "EXECUTIVE SUMMARY\nHigh-level compliance status for leadership"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    # Key Metrics section
    ws["A3"] = "KEY METRICS"
    ws["A3"].font = Font(bold=True, size=12)

    metrics = [
        ("Overall Completion Rate", "", "Target: ≥95%"),
        ("On-Time Completion Rate", "", "Target: ≥90%"),
        ("Average Assessment Score", "", "Target: ≥80%"),
        ("Phishing Click Rate", "", "Target: ≤5%"),
        ("Phishing Report Rate", "", "Target: ≥70%"),
        ("Active Remediation Cases", "", "Target: Minimized"),
    ]

    row = 4
    headers = ["Metric", "Current Value", "Target"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    row += 1
    for metric, value, target in metrics:
        ws.cell(row=row, column=1, value=metric).border = styles["border"]
        cell = ws.cell(row=row, column=2, value=value)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = styles["border"]
        cell.font = Font(bold=True, size=12)
        ws.cell(row=row, column=3, value=target).border = styles["border"]
        row += 1

    # Compliance Status
    row += 2
    ws[f"A{row}"] = "COMPLIANCE STATUS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    ws[f"A{row}"] = "Overall Status:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]

    dv_status = DataValidation(type="list", formula1='"Green - Compliant,Yellow - Minor Gaps,Red - Significant Gaps"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row}"])

    # Top 5 Risk Areas
    row += 3
    ws[f"A{row}"] = "TOP 5 RISK AREAS"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    headers = ["Priority", "Risk Area", "Current Status", "Action Required"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    for i in range(5):
        row += 1
        ws.cell(row=row, column=1, value=i + 1).border = styles["border"]
        for c in range(2, 5):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    # Key Achievements
    row += 3
    ws[f"A{row}"] = "KEY ACHIEVEMENTS THIS PERIOD"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    for i in range(3):
        ws[f"A{row}"] = f"{i + 1}."
        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40


def create_compliance_metrics_sheet(ws, styles):
    """Create Compliance_Metrics sheet."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "COMPLIANCE METRICS\nDetailed compliance breakdowns"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Category", 25),
        ("Total_Population", 15),
        ("Training_Required", 18),
        ("Completed", 12),
        ("Completion_Rate", 15),
        ("On_Time_Rate", 15),
        ("Average_Score", 15),
        ("Overdue_Count", 15),
        ("Remediation_Active", 18),
        ("Compliance_Status", 18),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_status = DataValidation(type="list", formula1='"Green,Yellow,Red"', allow_blank=False)
    ws.add_data_validation(dv_status)

    for r in range(4, 34):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]

        # Completion Rate formula
        ws.cell(row=r, column=5, value=f'=IF(C{r}=0,"N/A",ROUND(D{r}/C{r}*100,1)&"%")')
        ws.cell(row=r, column=5).font = Font(bold=True)

        dv_status.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "A4"


def create_effectiveness_metrics_sheet(ws, styles):
    """Create Effectiveness_Metrics sheet with pre-populated data."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "EFFECTIVENESS METRICS\nMeasure training impact on security behaviour"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Metric_Category", 18),
        ("Metric_Name", 30),
        ("Baseline_Value", 15),
        ("Current_Value", 15),
        ("Target_Value", 15),
        ("Trend_Direction", 15),
        ("Period_Comparison", 18),
        ("Data_Source", 30),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Pre-populate effectiveness metrics
    row += 1
    for category, metric, target, source in EFFECTIVENESS_METRICS:
        ws.cell(row=row, column=1, value=category).border = styles["border"]
        ws.cell(row=row, column=2, value=metric).border = styles["border"]
        for c in [3, 4]:
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        ws.cell(row=row, column=5, value=target).border = styles["border"]
        for c in [6, 7]:
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        ws.cell(row=row, column=8, value=source).border = styles["border"]
        ws.cell(row=row, column=9).fill = styles["input_cell"]["fill"]
        ws.cell(row=row, column=9).border = styles["border"]
        row += 1

    # Add empty rows
    for r in range(row, row + 10):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    dv_trend = DataValidation(type="list", formula1='"↑ Improving,→ Stable,↓ Declining,↑ Worsening"', allow_blank=False)
    ws.add_data_validation(dv_trend)
    for r in range(4, row + 10):
        dv_trend.add(ws.cell(row=r, column=6))

    ws.freeze_panes = "A4"


def create_trend_analysis_sheet(ws, styles):
    """Create Trend_Analysis sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "TREND ANALYSIS\nHistorical performance tracking"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Period", 15),
        ("Completion_Rate", 15),
        ("On_Time_Rate", 15),
        ("Average_Score", 15),
        ("Phishing_Click_Rate", 18),
        ("Phishing_Report_Rate", 18),
        ("Incident_Count", 15),
        ("Remediation_Count", 18),
        ("Compliance_Score", 15),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Pre-populate period labels
    periods = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    row += 1
    for period in periods:
        ws.cell(row=row, column=1, value=period).border = styles["border"]
        for c in range(2, 10):
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    ws.freeze_panes = "A4"


def create_risk_heatmap_sheet(ws, styles):
    """Create Risk_Heatmap sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "RISK HEATMAP\nVisual risk identification for training gaps"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Risk Scoring: 0=None, 1=Low, 2=Medium, 3=High"
    ws["A3"].font = Font(italic=True)

    headers = [
        ("Entity", 25),
        ("Overdue_Risk", 15),
        ("Assessment_Risk", 15),
        ("Simulation_Risk", 15),
        ("Incident_Risk", 15),
        ("Combined_Risk", 15),
        ("Priority_Rank", 12),
        ("Recommended_Action", 40),
    ]

    row = 5
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_risk = DataValidation(type="list", formula1='"0,1,2,3"', allow_blank=False)
    dv_combined = DataValidation(type="list", formula1='"Low,Medium,High,Critical"', allow_blank=False)
    ws.add_data_validation(dv_risk)
    ws.add_data_validation(dv_combined)

    for r in range(6, 26):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        for col in [2, 3, 4, 5]:
            dv_risk.add(ws.cell(row=r, column=col))
        dv_combined.add(ws.cell(row=r, column=6))

        # Combined Risk formula
        ws.cell(row=r, column=6, value=f'=IF(SUM(B{r}:E{r})>=8,"Critical",IF(SUM(B{r}:E{r})>=5,"High",IF(SUM(B{r}:E{r})>=2,"Medium","Low")))')

    ws.freeze_panes = "A6"


def create_audit_evidence_sheet(ws, styles):
    """Create Audit_Evidence sheet with pre-populated data."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "AUDIT EVIDENCE CHECKLIST\nEvidence status for audit readiness"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Evidence_Category", 15),
        ("Evidence_Item", 40),
        ("Required_For", 20),
        ("Status", 15),
        ("Location", 35),
        ("Last_Updated", 12),
        ("Responsible_Party", 20),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_status = DataValidation(type="list", formula1='"Available,Partial,Missing"', allow_blank=False)
    ws.add_data_validation(dv_status)

    row += 1
    for category, item, required, responsible in AUDIT_EVIDENCE:
        ws.cell(row=row, column=1, value=category).border = styles["border"]
        ws.cell(row=row, column=2, value=item).border = styles["border"]
        ws.cell(row=row, column=3, value=required).border = styles["border"]

        cell = ws.cell(row=row, column=4)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = styles["border"]
        dv_status.add(cell)

        for c in [5, 6]:
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

        ws.cell(row=row, column=7, value=responsible).border = styles["border"]

        cell = ws.cell(row=row, column=8)
        cell.fill = styles["input_cell"]["fill"]
        cell.border = styles["border"]

        row += 1

    ws.freeze_panes = "A4"


def create_management_review_input_sheet(ws, styles):
    """Create Management_Review_Input sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "MANAGEMENT REVIEW INPUT\nPrepared input for management review meetings"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    sections = [
        ("1. TRAINING PROGRAM STATUS SUMMARY", 5),
        ("2. EFFECTIVENESS ASSESSMENT", 5),
        ("3. NON-CONFORMITIES AND CORRECTIVE ACTIONS", 5),
        ("4. RESOURCE REQUIREMENTS", 4),
        ("5. RECOMMENDATIONS FOR IMPROVEMENT", 4),
        ("6. CHANGES AFFECTING TRAINING PROGRAM", 4),
    ]

    row = 3
    for section_title, lines in sections:
        ws[f"A{row}"] = section_title
        ws[f"A{row}"].font = Font(bold=True, size=11)
        row += 1

        for _ in range(lines):
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"].fill = styles["input_cell"]["fill"]
            ws[f"A{row}"].border = styles["border"]
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = 40
            row += 1

        row += 1

    ws.column_dimensions["A"].width = 100


def create_kpi_definitions_sheet(ws, styles):
    """Create KPI_Definitions sheet with pre-populated data."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "KPI DEFINITIONS\nReference for all metrics and KPIs"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("KPI_ID", 12),
        ("KPI_Name", 30),
        ("Definition", 50),
        ("Calculation_Method", 35),
        ("Target_Value", 15),
        ("Measurement_Frequency", 20),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row += 1
    for kpi_id, name, definition, calculation, target, frequency in KPI_DEFINITIONS:
        ws.cell(row=row, column=1, value=kpi_id).border = styles["border"]
        ws.cell(row=row, column=2, value=name).border = styles["border"]
        ws.cell(row=row, column=3, value=definition).border = styles["border"]
        ws.cell(row=row, column=4, value=calculation).border = styles["border"]
        ws.cell(row=row, column=5, value=target).border = styles["border"]
        ws.cell(row=row, column=6, value=frequency).border = styles["border"]
        row += 1

    ws.freeze_panes = "A4"


def create_data_sources_sheet(ws, styles):
    """Create Data_Sources sheet."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "DATA SOURCES\nDocument data inputs and refresh procedures"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 40

    headers = [
        ("Data_Element", 30),
        ("Source_System", 25),
        ("Source_Sheet", 25),
        ("Refresh_Frequency", 18),
        ("Last_Refresh", 12),
        ("Responsible_Party", 20),
        ("Validation_Method", 30),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Pre-populate common data sources
    data_sources = [
        ("Personnel List", "HRIS", "IMP-A.6.3.3 Personnel_Register", "Weekly"),
        ("Completion Records", "LMS", "IMP-A.6.3.3 Completion_Tracking", "Daily"),
        ("Assessment Scores", "LMS", "IMP-A.6.3.3 Assessment_Results", "Daily"),
        ("Simulation Results", "Phishing Platform", "IMP-A.6.3.3 Simulation_Results", "Per Campaign"),
        ("Remediation Status", "LMS/Manual", "IMP-A.6.3.3 Remediation_Tracking", "Weekly"),
        ("Incident Data", "SIEM/Ticketing", "Incident Management System", "Monthly"),
    ]

    row += 1
    for element, system, source, frequency in data_sources:
        ws.cell(row=row, column=1, value=element).border = styles["border"]
        ws.cell(row=row, column=2, value=system).border = styles["border"]
        ws.cell(row=row, column=3, value=source).border = styles["border"]
        ws.cell(row=row, column=4, value=frequency).border = styles["border"]
        for c in [5, 6, 7]:
            cell = ws.cell(row=row, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Add empty rows
    for r in range(row, row + 10):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]

    ws.freeze_panes = "A4"


def create_evidence_register_sheet(ws, styles):
    """Create Evidence_Register sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    headers = [
        ("Evidence_ID", 15),
        ("Evidence_Type", 22),
        ("Description", 40),
        ("Location", 40),
        ("Date_Collected", 15),
        ("Collected_By", 20),
        ("Status", 18),
        ("Notes", 40),
    ]

    row = 3
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    dv_status = DataValidation(type="list", formula1='"Verified,Pending,Requires update,Not available"', allow_blank=False)
    ws.add_data_validation(dv_status)

    for r in range(4, 54):
        ws.cell(row=r, column=1, value=f"EV-{r-3:03d}").font = Font(color="808080")
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
            cell.alignment = styles["input_cell"]["alignment"]
        dv_status.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "A4"


def create_dashboard_sheet(ws, styles):
    """Create Dashboard sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = "TRAINING COMPLIANCE DASHBOARD"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Compliance Summary (from Compliance_Metrics)"
    ws["A3"].font = Font(bold=True, size=12)

    metrics = [
        ("Total Completion Rate", "=AVERAGE(Compliance_Metrics!E4:E33)"),
        ("Total Overdue Count", "=SUM(Compliance_Metrics!H4:H33)"),
        ("Departments with Gaps", '=COUNTIF(Compliance_Metrics!J4:J33,"<>Green")'),
    ]

    row = 4
    for label, formula in metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF", size=14)
        row += 1

    row += 2
    ws[f"A{row}"] = "Effectiveness Summary (from Effectiveness_Metrics)"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    eff_metrics = [
        ("Metrics Meeting Target", '=COUNTIF(Effectiveness_Metrics!F4:F20,"*Improving*")+COUNTIF(Effectiveness_Metrics!F4:F20,"*Stable*")'),
        ("Metrics Below Target", '=COUNTIF(Effectiveness_Metrics!F4:F20,"*Declining*")+COUNTIF(Effectiveness_Metrics!F4:F20,"*Worsening*")'),
    ]

    for label, formula in eff_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    row += 2
    ws[f"A{row}"] = "Audit Readiness (from Audit_Evidence)"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    audit_metrics = [
        ("Evidence Available", '=COUNTIF(Audit_Evidence!D4:D20,"Available")'),
        ("Evidence Partial", '=COUNTIF(Audit_Evidence!D4:D20,"Partial")'),
        ("Evidence Missing", '=COUNTIF(Audit_Evidence!D4:D20,"Missing")'),
    ]

    for label, formula in audit_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    row += 2
    ws[f"A{row}"] = "Risk Summary (from Risk_Heatmap)"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    row += 1
    risk_metrics = [
        ("Critical Risk Areas", '=COUNTIF(Risk_Heatmap!F6:F25,"Critical")'),
        ("High Risk Areas", '=COUNTIF(Risk_Heatmap!F6:F25,"High")'),
        ("Medium Risk Areas", '=COUNTIF(Risk_Heatmap!F6:F25,"Medium")'),
    ]

    for label, formula in risk_metrics:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(bold=True, color="0000FF")
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20


def create_approval_signoff_sheet(ws, styles):
    """Create Approval_Sign_Off sheet."""
    ws.merge_cells("A1:E1")
    ws["A1"] = "QUARTERLY ATTESTATION AND SIGN-OFF"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 30

    row = 3
    ws[f"A{row}"] = "Reporting Period"
    ws[f"A{row}"].font = Font(bold=True, size=12)

    summary_fields = [
        ("Quarter/Period", ""),
        ("Report Date", ""),
        ("Overall Compliance Status", ""),
    ]

    row += 1
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = styles["input_cell"]["fill"]
        ws[f"B{row}"].border = styles["border"]
        row += 1

    dv_status = DataValidation(type="list", formula1='"Compliant,Minor Gaps,Significant Gaps"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add(ws[f"B{row-1}"])

    # Attestation sections
    sections = [
        ("PREPARED BY (Information Security Officer)", "4472C4"),
        ("REVIEWED BY (HR Director)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    for section_title, color in sections:
        row += 2
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = section_title
        ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

        fields = ["Name", "Date", "Signature"]
        row += 1
        for field in fields:
            ws[f"A{row}"] = field + ":"
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = styles["input_cell"]["fill"]
            ws[f"B{row}"].border = styles["border"]
            row += 1

    # Next Review
    row += 2
    ws[f"A{row}"] = "Next Review Date:"
    ws[f"A{row}"].font = Font(bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = styles["input_cell"]["fill"]
    ws[f"B{row}"].border = styles["border"]

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = setup_styles()

        logger.info("[1/13] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions"], styles)

        logger.info("[2/13] Creating Executive_Summary sheet...")
        create_executive_summary_sheet(wb["Executive_Summary"], styles)

        logger.info("[3/13] Creating Compliance_Metrics sheet...")
        create_compliance_metrics_sheet(wb["Compliance_Metrics"], styles)

        logger.info("[4/13] Creating Effectiveness_Metrics sheet...")
        create_effectiveness_metrics_sheet(wb["Effectiveness_Metrics"], styles)

        logger.info("[5/13] Creating Trend_Analysis sheet...")
        create_trend_analysis_sheet(wb["Trend_Analysis"], styles)

        logger.info("[6/13] Creating Risk_Heatmap sheet...")
        create_risk_heatmap_sheet(wb["Risk_Heatmap"], styles)

        logger.info("[7/13] Creating Audit_Evidence sheet...")
        create_audit_evidence_sheet(wb["Audit_Evidence"], styles)

        logger.info("[8/13] Creating Management_Review_Input sheet...")
        create_management_review_input_sheet(wb["Management_Review_Input"], styles)

        logger.info("[9/13] Creating KPI_Definitions sheet...")
        create_kpi_definitions_sheet(wb["KPI_Definitions"], styles)

        logger.info("[10/13] Creating Data_Sources sheet...")
        create_data_sources_sheet(wb["Data_Sources"], styles)

        logger.info("[11/13] Creating Evidence_Register sheet...")
        create_evidence_register_sheet(wb["Evidence_Register"], styles)

        logger.info("[12/13] Creating Dashboard sheet...")
        create_dashboard_sheet(wb["Dashboard"], styles)

        logger.info("[13/13] Creating Approval_Sign_Off sheet...")
        create_approval_signoff_sheet(wb["Approval_Sign_Off"], styles)

        wb.save(OUTPUT_FILENAME)

        logger.info("SUCCESS: %s", OUTPUT_FILENAME)
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error("Missing dependency: %s", e)
        return 1
    except Exception as e:
        logger.error("Unexpected error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation following ISMS-IMP-A.6.3.4 specification
# =============================================================================
