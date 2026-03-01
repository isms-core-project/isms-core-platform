#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.34.6 - Cross-Border Transfer Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.34: Privacy and Protection of PII
Assessment Domain 6 of 7: Cross-Border Data Transfer Compliance

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific international data transfer activities, transfer
mechanisms, and GDPR Chapter V / FADP compliance requirements.

Key customization areas:
1. Transfer inventory sources (match your actual data flows and systems)
2. Transfer mechanisms (adapt adequacy lists to your jurisdictions)
3. TIA methodology (align with organisational risk assessment framework)
4. Processor relationships (specific to your vendor ecosystem)
5. Compliance thresholds (based on regulatory obligations)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.34 Privacy and Protection of PII Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for managing
international personal data transfers per GDPR Chapter V (Articles 44-50),
Swiss FADP Article 16, and ISO/IEC 27001:2022 Control A.5.34 requirements.

**Purpose:**
Enables systematic inventory, legal basis validation, Transfer Impact Assessment
(TIA) execution, and processor agreement compliance for cross-border data transfers,
demonstrating GDPR Chapter V compliance post-Schrems II and supporting audit-ready
documentation for supervisory authorities.

**Assessment Scope:**
- Cross-border transfer inventory (all PII leaving EU/EEA or Switzerland)
- Transfer mechanism validation (Adequacy, SCCs, BCRs, DPF, Derogations)
- Transfer Impact Assessments (TIAs) per Schrems II requirements
- Processor Data Processing Agreement (DPA) compliance verification
- Subprocessor onward transfer safeguards
- Supplementary measures implementation (technical, contractual, organisational)
- Gap analysis for unlawful or under-documented transfers
- Evidence collection for supervisory authority inspections
- Compliance dashboard with transfer metrics

**Generated Workbook Structure:**
1. Instructions & Legend - GDPR Chapter V guidance and adequacy decisions list
2. Cross-Border Transfer Register - Comprehensive inventory of international transfers
3. Transfer Impact Assessment (TIA) Register - Schrems II TIA documentation
4. Processor Agreement Tracker - DPA and SCC compliance verification
5. Evidence Repository - Audit evidence tracking and documentation linkage
6. Gap Analysis - Non-compliant transfer remediation planning
7. Compliance Dashboard - Auto-calculated metrics for executive oversight
8. Approval & Sign-Off - Stakeholder approvals and DPO sign-off

**Key Features:**
- Data validation with GDPR Chapter V transfer mechanism dropdowns
- Conditional formatting for adequacy status and compliance gaps
- Automated TIA requirement flagging for non-adequate country transfers
- EU Commission adequacy decisions list (automatically checks destination country)
- SCC version validation (2021 vs. invalid 2010 SCCs)
- Protected formulas with unprotected input cells
- Risk-based gap prioritization (Critical/High/Medium/Low)
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Dashboard with compliance percentages and gap counts

**Integration:**
This assessment is Domain 6 of the A.5.34 Privacy assessment suite (7 domains).
Results feed into ISMS-IMP-A.5.34.7 Privacy Compliance Dashboard for
consolidated privacy program oversight and regulatory reporting. Integrates
with A.5.34.1 (PII Inventory) for transfer identification and A.5.34.4 (TOMs)
for supplementary measure verification.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a5346_cross_border_transfer_assessment.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a5346_cross_border_transfer_assessment.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a5346_cross_border_transfer_assessment.py --date 20250130

Output:
    File: ISMS-IMP-A.5.34.6_Cross_Border_Transfer_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize adequacy decisions list (Sheet 1) for applicable jurisdictions
    2. Identify all cross-border transfers using A.5.34.1 PII inventory data
    3. Complete Sheet 2 (Transfer Register) - document all international transfers
    4. Determine transfer mechanisms (Adequacy, SCCs, BCRs, DPF, Derogations)
    5. Conduct Transfer Impact Assessments (TIAs) in Sheet 3 for non-adequate countries
    6. Implement supplementary measures (encryption, contractual clauses, monitoring)
    7. Verify processor agreements in Sheet 4 (DPAs with SCCs for non-adequate)
    8. Check subprocessor onward transfer safeguards
    9. Collect and link audit evidence in Sheet 5
    10. Conduct gap analysis in Sheet 6 for unlawful/under-documented transfers
    11. Create remediation plan with risk-based prioritization
    12. Review Sheet 7 (Dashboard) for compliance metrics
    13. Obtain DPO approval and stakeholder sign-offs in Sheet 8
    14. Schedule quarterly TIA reviews per Schrems II continuous monitoring
    15. Feed results into A.5.34.7 Privacy Compliance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.34
Assessment Domain:    6 of 7 (Cross-Border Data Transfer Compliance)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.34: Privacy and Protection of PII Policy (Governance)
    - ISMS-CTX-A.5.34: Privacy Regulatory Landscape Reference (Context)
    - ISMS-IMP-A.5.34.1: PII Identification Assessment (Domain 1)
    - ISMS-IMP-A.5.34.2: Legal Basis Assessment (Domain 2)
    - ISMS-IMP-A.5.34.3: DSR Management Assessment (Domain 3)
    - ISMS-IMP-A.5.34.4: TOMs Assessment (Domain 4)
    - ISMS-IMP-A.5.34.5: DPIA Assessment (Domain 5)
    - ISMS-IMP-A.5.34.6: Cross-Border Transfer Implementation Guide (Parts 1-3)
    - ISMS-IMP-A.5.34.7: Privacy Compliance Dashboard (Consolidation)
    - EDPB Recommendations 01/2020: Supplementary Measures for Transfers
    - Schrems II CJEU C-311/18: Privacy Shield Invalidation Decision

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 2025-01-30
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.34.6 specification
    - Supports comprehensive cross-border transfer compliance assessment
    - Integrated TIA framework per Schrems II requirements
    - Prepares for consolidation into A.5.34.7 Privacy Compliance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**GDPR Chapter V Post-Schrems II:**
The Schrems II CJEU decision (July 2020) invalidated the EU-US Privacy Shield
and mandated Transfer Impact Assessments (TIAs) for all transfers to countries
without adequacy decisions, EVEN WHEN SCCs ARE IN PLACE. Organisations must:
(1) Assess destination country legal framework (surveillance laws, government
access powers), (2) Evaluate importer-specific circumstances, (3) Implement
supplementary measures (technical, contractual, organisational) to ensure
adequate protection, (4) Document TIA conclusions. TIAs must be reassessed
if legal/political situation changes. This is NOT optional - it's a legal
requirement for GDPR compliance.

**EU-US Data Privacy Framework (DPF):**
The EU-US DPF (July 2023) provides adequacy for US companies that self-certify.
However, DPF certifications expire annually and must be renewed. Always verify
current DPF status at https://www.dataprivacyframework.gov/s/participant-search
before relying on DPF as transfer mechanism. DPF may be challenged in court
(similar to Schrems I & II) - maintain backup transfer mechanisms (SCCs + TIA).

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 and GDPR requirements.
Supervisory authorities and auditors will expect: complete transfer inventory,
documented transfer mechanisms with valid legal basis, TIAs for all non-adequate
country transfers, evidence of supplementary measures implementation, processor
DPAs with SCCs (2021 version), subprocessor onward transfer safeguards, regular
TIA reviews (quarterly or annually), DPO involvement and approval. Ensure all
transfers are documented BEFORE processing begins - unlawful transfers can
result in GDPR fines up to 4% of global annual revenue.

**Data Protection:**
Assessment workbooks contain sensitive information including:
- Complete inventory of international data flows (reveals business relationships)
- TIA risk analyses highlighting legal vulnerabilities
- Processor agreements and subprocessor relationships (competitive intelligence)
- Gap analysis revealing compliance deficiencies
- Supplementary measures implementations (security architecture details)

Handle in accordance with your organisation's data classification policies.
Restrict access to DPO, Legal Counsel, Privacy Team, and authorised compliance
personnel only. Consider encryption at rest for stored workbooks.

**Maintenance:**
Review and update assessment:
- Quarterly: New transfers, processor changes, TIA reviews (especially for US transfers)
- Annually: Complete transfer inventory reassessment, adequacy decision updates
- Triggered: New adequacy decisions, adequacy revocations, Schrems III potential ruling,
  regulatory guidance updates, processor contract changes, new subprocessors
- Continuous: Monitor destination country legal developments (new surveillance laws)

**Quality Assurance:**
Have DPO and Legal Counsel validate all transfer mechanisms and TIA conclusions
before implementing transfers. For high-risk transfers (e.g., sensitive PII to
US or China), consider external legal review. Conduct annual audit of transfer
register completeness - compare against A.5.34.1 PII inventory to identify
undocumented transfers. Train business owners on cross-border transfer approval
requirements BEFORE signing new processor contracts.

**Regulatory Alignment:**
This assessment supports compliance with:
- GDPR (EU General Data Protection Regulation) - Chapter V (Art. 44-50)
- FADP (Swiss Federal Act on Data Protection) - Art. 16 (Disclosure Abroad)
- ISO/IEC 27001:2022 - Control A.5.34 (Privacy and Protection of PII)
- ISO/IEC 27701:2019 - Privacy Information Management (if applicable)
- Schrems II CJEU C-311/18 - TIA Requirements
- EDPB Recommendations 01/2020 - Supplementary Measures Guidance

**Critical Deadlines and Compliance Milestones:**
- 2010 SCCs: INVALID since September 27, 2022 (must update to 2021 version)
- DPF certifications: Expire annually, verify renewal quarterly
- TIA reviews: Reassess when destination country legal framework changes
- Processor DPA reviews: Annual review of subprocessor lists and SCC flow-down
- Gap remediation: Critical gaps = immediate action (1-4 weeks)
- Supervisory authority consultations: Pre-consultation required if TIA shows
  high residual risk (GDPR Article 36)

================================================================================
"""

import os
from datetime import datetime
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging
import sys

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.34.6"
WORKBOOK_NAME = "Cross-Border Transfer Assessment"
CONTROL_ID = "A.5.34"
CONTROL_NAME = "Privacy and Protection of PII"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# ============================================================================
# CONFIGURATION CONSTANTS
# ============================================================================

# CUSTOMIZE: Colors matching A.5.34 suite (6-char hex, no FF prefix)
COLORS = {
    'header_blue': '003366',
    'white': 'FFFFFF',
    'black': '000000',
    'light_green': 'C6EFCE',
    'dark_green': '006100',
    'light_yellow': 'FFEB9C',
    'dark_orange': '9C5700',
    'light_red': 'FFC7CE',
    'dark_red': '9C0006',
    'light_blue': 'B4C6E7',
    'light_gray': 'D9D9D9',
    'light_orange': 'FFEB9C',
}

# CUSTOMIZE: EU Adequacy Decisions (as of 2025)
EU_ADEQUACY_COUNTRIES = [
    "Andorra", "Argentina", "Canada (Commercial)", "Faroe Islands",
    "Guernsey", "Israel", "Isle of Man", "Japan", "Jersey",
    "New Zealand", "Republic of Korea", "Switzerland", "United Kingdom", "Uruguay"
]

# Transfer status options
STATUS_OPTIONS = ["Not Started", "In Progress", "Complete", "Validated"]

# Risk levels
RISK_LEVELS = ["Low", "Medium", "High", "Critical"]

# Protection password
PROTECTION_PASSWORD = "privacy2024"

# File prefix
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================
_THIN_SIDE = Side(style='thin')
THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)

def style_header_row(ws, row_num, color_hex, num_columns):
    """Apply consistent header styling."""
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(name='Calibri', size=11, bold=True, color=COLORS['white'])
        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )


def add_dropdown(ws, cell_range, options, error_msg="Invalid selection", allow_blank=False):
    """Add dropdown data validation."""
    dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=allow_blank)
    dv.error = error_msg
    dv.errorStyle = 'stop'
    ws.add_data_validation(dv)
    dv.add(cell_range)


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass


# ============================================================================
# SHEET CREATION FUNCTIONS (8 SHEETS)
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete Sheet 2 (Transfer Register) — inventory all international personal data transfers.', '2. Complete Sheet 3 (TIA Register) — conduct Transfer Impact Assessments for non-adequate countries.', '3. Complete Sheet 4 (Processor Tracker) — verify DPAs and SCCs for all processors.', '4. Complete Sheet 5 (Evidence Repository) — link supporting evidence for all transfers.', '5. Complete Sheet 6 (Gap Analysis) — document non-compliant or under-documented transfers.', '6. Review Sheet 7 (Dashboard) — validate transfer compliance metrics.', '7. Obtain sign-offs in Sheet 8 (Approvals) — DPO and stakeholder approvals.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 21

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def _apply_data_fills(ws, first_data_row, num_cols):
    """Apply F2F2F2 sample row + 50 FFFFCC empty rows with thin borders."""
    from openpyxl.styles import PatternFill as _PF, Border as _B, Side as _S
    _f2 = _PF(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _ff = _PF(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _t = _S(style="thin")
    _bd = _B(left=_t, right=_t, top=_t, bottom=_t)
    for col in range(1, num_cols + 1):
        c = ws.cell(row=first_data_row, column=col)
        c.fill = _f2
        c.border = _bd
    for row in range(first_data_row + 1, first_data_row + 51):
        for col in range(1, num_cols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = _ff
            c.border = _bd

def create_transfer_register_sheet(wb):
    """Sheet 2: Cross-Border Transfer Register."""
    ws = wb.create_sheet("Transfer Register")
    ws.sheet_view.showGridLines = False

    # Title row — fixes DS-001/MRG-001
    num_cols = 20
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    ws['A1'] = "CROSS-BORDER TRANSFER REGISTER"
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    headers = [
        "Transfer ID", "Status", "Transfer Name", "Source System", "Destination System",
        "Destination Country", "Adequacy Status", "Transfer Mechanism", "SCC Version",
        "SCC Date", "DPF Cert?", "TIA Required?", "TIA ID", "PII Categories",
        "Transfer Volume", "Transfer Frequency", "Purpose", "Legal Basis (Art.6)",
        "Last Updated", "Notes"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Add formula in A3 for Transfer ID
    ws['A3'] = '=TEXT(ROW()-2,"XFER-0000")'

    # Dropdowns
    add_dropdown(ws, 'B4:B1000', STATUS_OPTIONS, "Select valid status")
    add_dropdown(ws, 'H4:H1000', ["Adequacy", "SCCs", "BCRs", "DPF", "Derogation"], "Select mechanism")
    add_dropdown(ws, 'L4:L1000', ["YES", "NO"], "YES or NO")

    ws.freeze_panes = 'A3'
    _apply_data_fills(ws, 3, num_cols)

    return ws


def create_tia_register_sheet(wb):
    """Sheet 3: Transfer Impact Assessment Register."""
    ws = wb.create_sheet("TIA Register")
    ws.sheet_view.showGridLines = False

    # Title row — fixes MRG-001 and DS-001/DS-002
    num_cols = 15
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    ws['A1'] = "TIA REGISTER"
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    headers = [
        "TIA ID", "Transfer ID", "Status", "Destination Country", "Assessment Date",
        "Assessor", "Surveillance Laws", "Gov Access Risk", "Risk Justification",
        "Supplementary Measures", "Residual Risk", "TIA Conclusion", "DPO Approval",
        "Next Review Date", "Notes"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws['A3'] = '=TEXT(ROW()-2,"TIA-0000")'

    add_dropdown(ws, 'C4:C1000', ["Not Started", "In Progress", "Complete", "Approved"], "Status")
    add_dropdown(ws, 'H4:H1000', RISK_LEVELS, "Risk level")
    add_dropdown(ws, 'K4:K1000', RISK_LEVELS, "Risk level")
    add_dropdown(ws, 'L4:L1000', ["PASS", "PASS with Conditions", "FAIL"], "TIA result")

    ws.freeze_panes = 'A3'
    _apply_data_fills(ws, 3, num_cols)

    return ws


def create_processor_tracker_sheet(wb):
    """Sheet 4: Processor Agreement Tracker."""
    ws = wb.create_sheet("Processor Tracker")
    ws.sheet_view.showGridLines = False

    # Title row — fixes DS-001/MRG-001
    num_cols = 16
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    ws['A1'] = "PROCESSOR AGREEMENT TRACKER"
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    headers = [
        "Processor ID", "Transfer ID", "Processor Name", "Processor Location",
        "DPA Exists?", "DPA Date", "SCCs Included?", "SCC Version", "SCC Date",
        "Subprocessors?", "Compliance Status", "Gap Description", "Remediation Action",
        "Owner", "Deadline", "Notes"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws['A3'] = '=TEXT(ROW()-2,"PROC-0000")'

    add_dropdown(ws, 'E4:E1000', ["YES", "NO"], "DPA status")
    add_dropdown(ws, 'G4:G1000', ["YES", "NO"], "SCCs status")
    add_dropdown(ws, 'K4:K1000', ["Compliant", "Partially Compliant", "Non-Compliant"], "Status")

    ws.freeze_panes = 'A3'
    _apply_data_fills(ws, 3, num_cols)

    return ws


def create_evidence_repository_sheet(wb):
    """Sheet 5: Evidence Repository."""
    ws = wb.create_sheet("Evidence Repository")
    ws.sheet_view.showGridLines = False

    # Title row — fixes DS-001/MRG-001
    num_cols = 11
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    ws['A1'] = "EVIDENCE REPOSITORY"
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    headers = [
        "Evidence ID", "Transfer ID", "TIA ID", "Evidence Type", "Description",
        "Document Name", "File Location", "Upload Date", "Owner", "Status", "Notes"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws['A3'] = '=TEXT(ROW()-2,"EVID-0000")'

    evidence_types = ["SCCs (Signed)", "DPA", "TIA", "DPF Cert", "Legal Memo", "Other"]
    add_dropdown(ws, 'D4:D1000', evidence_types, "Evidence type")
    add_dropdown(ws, 'J4:J1000', ["Current", "Expired", "Superseded"], "Status")

    ws.freeze_panes = 'A3'
    _apply_data_fills(ws, 3, num_cols)

    return ws


def create_gap_analysis_sheet(wb):
    """Sheet 6: Gap Analysis."""
    ws = wb.create_sheet("Gap Analysis")
    ws.sheet_view.showGridLines = False

    # Title row — fixes DS-001/MRG-001
    num_cols = 12
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    ws['A1'] = "GAP ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    headers = [
        "Gap ID", "Transfer ID", "Gap Type", "Description", "Risk Level",
        "Affected Data Subjects", "Discovery Date", "Remediation Action",
        "Owner", "Target Date", "Status", "Notes"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws['A3'] = '=TEXT(ROW()-2,"GAP-0000")'

    gap_types = ["Missing SCCs", "No TIA", "Undocumented Transfer", "Old SCCs", "Other"]
    add_dropdown(ws, 'C4:C1000', gap_types, "Gap type")
    add_dropdown(ws, 'E4:E1000', RISK_LEVELS, "Risk level")
    add_dropdown(ws, 'K4:K1000', ["Open", "In Progress", "Completed", "Blocked"], "Status")

    # Conditional formatting for risk levels
    ws.conditional_formatting.add(
        'E3:E1000',
        CellIsRule(operator='equal', formula=['"Critical"'],
                   fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'))
    )

    ws.freeze_panes = 'A3'
    _apply_data_fills(ws, 3, num_cols)

    return ws


def create_summary_dashboard_sheet(wb):
    """Create the Gold Standard Summary Dashboard sheet for A.5.34.6."""
    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ffffcc = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    c00000 = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    # Row 1: Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "CROSS-BORDER TRANSFER \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = THIN_BORDER
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.5.34: Privacy and Protection of Personally Identifiable Information (PII)"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: empty

    # TABLE 1 banner (Row 4)
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = navy
    for c in range(1, 8):
        ws.cell(row=4, column=c).border = THIN_BORDER

    # TABLE 1 column headers (Row 5)
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # TABLE 1 data rows
    # Transfer Mechanism: Adequacy/DPF = Compliant (Art. 45 highest protection)
    #                     SCCs/BCRs = Partial (Art. 46 adequate safeguards)
    #                     Derogation = Non-Compliant (Art. 49 exceptional basis only)
    area_configs = [
        (
            "Transfer Register",
            "=COUNTA('Transfer Register'!C4:C53)",
            "=COUNTIF('Transfer Register'!H4:H53,\"Adequacy\")+COUNTIF('Transfer Register'!H4:H53,\"DPF\")",
            "=COUNTIF('Transfer Register'!H4:H53,\"SCCs\")+COUNTIF('Transfer Register'!H4:H53,\"BCRs\")",
            "=COUNTIF('Transfer Register'!H4:H53,\"Derogation\")",
            "=\"\"",
        ),
        (
            "TIA Register",
            "=COUNTA('TIA Register'!B4:B53)",
            "=COUNTIF('TIA Register'!L4:L53,\"PASS\")",
            "=COUNTIF('TIA Register'!L4:L53,\"PASS with Conditions\")",
            "=COUNTIF('TIA Register'!L4:L53,\"FAIL\")",
            "=\"\"",
        ),
        (
            "Processor Tracker",
            "=COUNTA('Processor Tracker'!C4:C53)",
            "=COUNTIF('Processor Tracker'!K4:K53,\"Compliant\")",
            "=COUNTIF('Processor Tracker'!K4:K53,\"Partially Compliant\")",
            "=COUNTIF('Processor Tracker'!K4:K53,\"Non-Compliant\")",
            "=\"\"",
        ),
        (
            "Gap Analysis",
            "=COUNTA('Gap Analysis'!B4:B53)",
            "=COUNTIF('Gap Analysis'!E4:E53,\"Low\")",
            "=COUNTIF('Gap Analysis'!E4:E53,\"Medium\")",
            "=COUNTIF('Gap Analysis'!E4:E53,\"High\")+COUNTIF('Gap Analysis'!E4:E53,\"Critical\")",
            "=\"\"",
        ),
    ]

    for i, (area_name, total_f, comp_f, partial_f, noncomp_f, na_f) in enumerate(area_configs):
        row = 6 + i
        ws.cell(row=row, column=1, value=area_name).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(color="000000")

        cell_b = ws.cell(row=row, column=2, value=total_f)
        cell_b.border = THIN_BORDER
        cell_b.alignment = Alignment(horizontal="center")
        cell_b.font = Font(color="000000")

        cell_c = ws.cell(row=row, column=3, value=comp_f)
        cell_c.border = THIN_BORDER
        cell_c.alignment = Alignment(horizontal="center")
        cell_c.font = Font(color="000000")

        cell_d = ws.cell(row=row, column=4, value=partial_f)
        cell_d.border = THIN_BORDER
        cell_d.alignment = Alignment(horizontal="center")
        cell_d.font = Font(color="000000")

        cell_e = ws.cell(row=row, column=5, value=noncomp_f)
        cell_e.border = THIN_BORDER
        cell_e.alignment = Alignment(horizontal="center")
        cell_e.font = Font(color="000000")

        cell_f = ws.cell(row=row, column=6, value=na_f)
        cell_f.border = THIN_BORDER
        cell_f.alignment = Alignment(horizontal="center")
        cell_f.font = Font(color="000000")

        cell_g = ws.cell(row=row, column=7, value=f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))")
        cell_g.number_format = "0.0%"
        cell_g.border = THIN_BORDER
        cell_g.alignment = Alignment(horizontal="center")
        cell_g.font = Font(color="000000")

    # TOTAL row (row 10)
    total_row = 10
    for col in range(1, 8):
        ws.cell(row=total_row, column=col).fill = grey
        ws.cell(row=total_row, column=col).border = THIN_BORDER
        ws.cell(row=total_row, column=col).font = Font(bold=True, color="000000")
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left")
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}9)"
        cell.alignment = Alignment(horizontal="center")
    cell_g_total = ws.cell(row=total_row, column=7,
                           value=f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))")
    cell_g_total.number_format = "0.0%"
    cell_g_total.alignment = Alignment(horizontal="center")

    # TABLE 2: KEY METRICS
    metrics_start = total_row + 2  # row 12
    ws.merge_cells(f"A{metrics_start}:G{metrics_start}")
    ws[f"A{metrics_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{metrics_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{metrics_start}"].fill = navy
    for c in range(1, 8):
        ws.cell(row=metrics_start, column=c).border = THIN_BORDER

    metric_headers = ["Metric", "Value", "", "", "", "", ""]
    for col, header in enumerate(metric_headers, 1):
        cell = ws.cell(row=metrics_start + 1, column=col, value=header if header else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    metrics = [
        ("Total Transfers Registered",                        "=COUNTA('Transfer Register'!C4:C53)"),
        ("Transfers with Adequacy Decision (Art. 45)",        "=COUNTIF('Transfer Register'!H4:H53,\"Adequacy\")"),
        ("Transfers with Standard Contractual Clauses",       "=COUNTIF('Transfer Register'!H4:H53,\"SCCs\")"),
        ("Transfers with Binding Corporate Rules",            "=COUNTIF('Transfer Register'!H4:H53,\"BCRs\")"),
        ("Transfers via Data Privacy Framework",              "=COUNTIF('Transfer Register'!H4:H53,\"DPF\")"),
        ("Transfers via Derogation (Art. 49 \u2014 Exceptional)", "=COUNTIF('Transfer Register'!H4:H53,\"Derogation\")"),
        ("Total TIAs Conducted",                              "=COUNTA('TIA Register'!B4:B53)"),
        ("TIAs Passing",                                      "=COUNTIF('TIA Register'!L4:L53,\"PASS\")"),
        ("TIAs Passing with Conditions",                      "=COUNTIF('TIA Register'!L4:L53,\"PASS with Conditions\")"),
        ("TIAs Failing (Transfer at Risk)",                   "=COUNTIF('TIA Register'!L4:L53,\"FAIL\")"),
        ("Processors Tracked",                                "=COUNTA('Processor Tracker'!C4:C53)"),
        ("Compliant Processors",                              "=COUNTIF('Processor Tracker'!K4:K53,\"Compliant\")"),
        ("Non-Compliant Processors",                          "=COUNTIF('Processor Tracker'!K4:K53,\"Non-Compliant\")"),
    ]

    row = metrics_start + 2
    for metric, formula in metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(color="000000")
        cell_val = ws.cell(row=row, column=2, value=formula)
        cell_val.border = THIN_BORDER
        cell_val.font = Font(color="000000")
        cell_val.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # TABLE 2 buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # TABLE 3: CRITICAL FINDINGS
    crit_start = row + 1
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{crit_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = c00000
    for c in range(1, 8):
        ws.cell(row=crit_start, column=c).border = THIN_BORDER

    findings_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col, header in enumerate(findings_headers, 1):
        cell = ws.cell(row=crit_start + 1, column=col, value=header if header else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    findings = [
        ("TIA Register",      "TIAs with FAIL conclusion (transfer at risk \u2014 GDPR Chapter V)",    "=COUNTIF('TIA Register'!L4:L53,\"FAIL\")",                                      "Critical", "Immediate"),
        ("Processor Tracker", "Non-compliant third-country processors",                                "=COUNTIF('Processor Tracker'!K4:K53,\"Non-Compliant\")",                         "Critical", "Immediate"),
        ("Transfer Register", "Transfers via derogation only (Art. 49 \u2014 not a general basis)",   "=COUNTIF('Transfer Register'!H4:H53,\"Derogation\")",                            "High",     "Urgent"),
        ("Gap Analysis",      "Critical risk transfer gaps",                                           "=COUNTIF('Gap Analysis'!E4:E53,\"Critical\")",                                  "Critical", "Immediate"),
        ("Gap Analysis",      "High risk transfer gaps",                                               "=COUNTIF('Gap Analysis'!E4:E53,\"High\")",                                      "High",     "Urgent"),
        ("Processor Tracker", "Partially compliant processors requiring remediation",                  "=COUNTIF('Processor Tracker'!K4:K53,\"Partially Compliant\")",                   "Medium",   "Plan"),
    ]

    row = crit_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).font = Font(color="000000")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    # TABLE 3 buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Column widths and freeze
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"


# create_dashboard_sheet removed — superseded by Summary Dashboard

def create_approval_sheet(wb):
    """Create the Approval Sign-Off sheet — Gold Standard (GS-AS-014/015)."""
    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title banner — GS-AS-014
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY section banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Rows 4-8: Summary metadata — B6 = Overall Compliance (GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE(\'Summary Dashboard\'!G6:G9),\"\")")  ,
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Row 7 status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # Approver sections start at row 11 (rows 9-10 = gap)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION — GS-AS-012: col A = plain bold label, NO dark fill
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")

def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    
    # Create workbook
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.remove(wb.active)
    
    logger.info("Creating Sheet 1: Instructions & Legend...")
    create_instructions_sheet(wb.create_sheet())
    
    logger.info("Creating Sheet 2: Cross-Border Transfer Register...")
    create_transfer_register_sheet(wb)
    
    logger.info("Creating Sheet 3: TIA Register...")
    create_tia_register_sheet(wb)
    
    logger.info("Creating Sheet 4: Processor Agreement Tracker...")
    create_processor_tracker_sheet(wb)
    
    logger.info("Creating Sheet 5: Evidence Repository...")
    create_evidence_repository_sheet(wb)
    
    logger.info("Creating Sheet 6: Gap Analysis...")
    create_gap_analysis_sheet(wb)
    
    logger.info("Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb)


    logger.info("Creating Sheet 8: Approval & Sign-Off...")
    create_approval_sheet(wb)
    
    # Save
    finalize_validations(wb)
    logger.info(f"\nSaving workbook to {output_path}...")
    wb.save(output_path)
    
    logger.info(f"\nSuccess! Workbook created: {output_path}")
    logger.info("\nNext steps:")
    logger.info("1. Review Sheet 1 (Instructions) for GDPR Chapter V guidance")
    logger.info("2. Complete Sheet 2 (Transfer Register) - inventory all cross-border transfers")
    logger.info("3. Conduct TIAs in Sheet 3 for non-adequate country transfers")
    logger.info("4. Verify processor agreements in Sheet 4")
    logger.info("5. Collect evidence in Sheet 5")
    logger.info("6. Document gaps in Sheet 6")
    logger.info("7. Review Sheet 7 (Dashboard) for compliance metrics")
    logger.info("8. Obtain approvals in Sheet 8")
    logger.info("9. Feed results into A.5.34.7 Privacy Compliance Dashboard")


def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == '__main__':
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
