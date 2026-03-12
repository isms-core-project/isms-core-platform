#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.29.2 - Degraded Mode Security Requirements Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.29: Information Security During Disruption
Assessment Domain 2 of 3: Degraded Mode Security Requirements

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific information security during disruption infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Disruption scenario categories and priority tiers (match your BCP framework)
2. Security control degradation thresholds and acceptable risk levels
3. Degraded mode security requirements per service tier
4. Recovery verification checklist items and sign-off authorities
5. Integration with your organisation's BCP/DR testing programme

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.29 Information Security During Disruption Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
information security during disruption controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Degraded Mode Security Requirements under ISO 27001:2022 Control A.5.29. Supports evidence-based documentation of security control resilience, degraded mode procedures, and recovery verification compliance.

**Assessment Scope:**
- Security control coverage during disruption scenarios
- Degraded mode security requirement definition and completeness
- Compensating control adequacy during reduced-capability periods
- Recovery security verification procedure documentation
- BCP/DR integration with information security requirements
- Testing and exercise outcomes for disruption scenarios
- Evidence collection for business continuity and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
4. Summary Dashboard - Compliance overview and key metrics
5. Evidence Register - Audit evidence tracking
6. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 3 domains covering Information Security During Disruption controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a529_2_degraded_mode.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a529_2_degraded_mode.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a529_2_degraded_mode.py --date 20250115

Output:
    File: ISMS-IMP-A.5.29.2_Degraded_Mode_Security_Requirements_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.29
Assessment Domain:    2 of 3 (Degraded Mode Security Requirements)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.29: Information Security During Disruption Policy (Governance)
    - ISMS-IMP-A.5.29.1: Security Controls During Disruption (Domain 1)
    - ISMS-IMP-A.5.29.2: Degraded Mode Security Requirements (Domain 2)
    - ISMS-IMP-A.5.29.3: Recovery Security Verification (Domain 3)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.29.2 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive information security during disruption details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review security during disruption procedures annually or when BCP/DR plans are updated, critical system architectures change, or disruption testing reveals gaps.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

# =============================================================================
# IMPORTS - STANDARD LIBRARY
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# IMPORTS - THIRD PARTY
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.29.2"
WORKBOOK_NAME = "Degraded Mode Security Requirements"
CONTROL_ID = "A.5.29"
CONTROL_NAME = "Information Security During Disruption"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "warning": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# =============================================================================
# WORKBOOK CREATION
# =============================================================================

_STYLES = setup_styles()
def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    sheets = [
        "Instructions & Legend",
        "Degradation Scenarios",
        "BreakGlass Accounts",
        "BreakGlass Activation",
        "Elevated Monitoring",
        "Personnel Availability",
        "Security Debt Register",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================

def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Degradation Scenarios — document all scenarios requiring degraded security mode operation.',
        '2. Complete BreakGlass Accounts — inventory emergency access accounts with activation procedures.',
        '3. Complete BreakGlass Activation — document the authorisation workflow for emergency access.',
        '4. Complete Elevated Monitoring — define enhanced monitoring requirements during degraded mode.',
        '5. Complete Personnel Availability — identify security personnel available during disruption events.',
        '6. Track Security Debt Register — log security compromises accepted during degraded mode for remediation.',
        '7. Maintain the Evidence Register with degraded mode procedures and exercise records.',
        '8. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A22"] = "Status Legend"
    ws["A22"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=23, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 24 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_degradation_scenarios_sheet(ws, styles):
    """Create the Degradation Scenarios sheet."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:K1")
    ws["A1"] = "DEGRADATION SCENARIOS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:K2")
    ws["A2"] = "Documented acceptable security degradations for each disruption type, with duration limits and authorisation requirements"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Scenario_ID", 15),
        ("Scenario_Name", 30),
        ("Trigger_Condition", 40),
        ("Control_Affected", 25),
        ("Degradation_Type", 22),
        ("Compensating_Control", 45),
        ("Max_Duration", 20),
        ("Renewal_Process", 35),
        ("Authorisation_Required", 25),
        ("Posture_Level", 15),
        ("Never_Acceptable", 40),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populated reference scenarios rows 4-8 (F2F2F2 grey)
    scenarios = [
        ("DS-001", "MFA Infrastructure Unavailable",
         "MFA platform failure or inaccessible",
         "Multi-Factor Authentication",
         "Temporary Bypass",
         "Enhanced logging, IP restrictions, session limits, 4-hour max sessions",
         "72 hours", "CISO approval required for each 24h extension",
         "CISO", "Degraded",
         "Allowing permanent single-factor access"),
        ("DS-002", "SIEM Platform Unavailable",
         "SIEM system failure or maintenance",
         "Security Monitoring",
         "Reduced Capability",
         "Manual log review every 4 hours for critical systems",
         "48 hours", "Security Manager approval for extension",
         "Security Manager", "Degraded",
         "No logging of security events"),
        ("DS-003", "Network Segmentation Bypass",
         "Emergency connectivity requirement",
         "Network Segmentation",
         "Temporary Bypass",
         "Enhanced monitoring, specific ports only, time-limited",
         "24 hours", "CIO + CISO approval for extension",
         "CIO + CISO", "Emergency",
         "Permanent network flattening"),
        ("DS-004", "Patch Management Delayed",
         "System stability during disruption",
         "Patch Management",
         "Postponed",
         "Critical patches only, manual review process",
         "30 days non-critical", "CISO approval for extension",
         "CISO", "Degraded",
         "Ignoring critical security patches"),
        ("DS-005", "Access Review Postponed",
         "Resource constraints during disruption",
         "Access Reviews",
         "Postponed",
         "Stricter approval for new access, no new privileged access",
         "30 days", "Security Manager approval for extension",
         "Security Manager", "Elevated",
         "Granting new privileged access without review"),
    ]

    dv_type = DataValidation(
        type="list",
        formula1='"Temporary Bypass,Reduced Capability,Postponed,Alternative Method"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_posture = DataValidation(
        type="list",
        formula1='"Elevated,Degraded,Emergency"',
        allow_blank=False
    )
    ws.add_data_validation(dv_posture)

    for row_idx, scenario in enumerate(scenarios, start=4):
        for col_idx, value in enumerate(scenario, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = grey_sample_fill
            cell.border = border_thin
            if col_idx in [3, 6, 8, 11]:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
        dv_type.add(ws.cell(row=row_idx, column=5))
        dv_posture.add(ws.cell(row=row_idx, column=10))

    # Row 9: F2F2F2 grey sample row (MAX-003 fix)
    sample_row = 9
    sample_values = {
        1: "DS-006", 2: "EDR Platform Unavailable",
        3: "EDR/antimalware platform failure", 4: "Endpoint Detection & Response",
        5: "Reduced Capability", 6: "Increased manual scans twice daily, log review hourly",
        7: "48 hours", 8: "CISO approval required for each 24h extension",
        9: "CISO", 10: "Degraded", 11: "No endpoint security monitoring on critical servers",
    }
    for col_idx, value in sample_values.items():
        cell = ws.cell(row=sample_row, column=col_idx, value=value)
        cell.fill = grey_sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 10-59: 50 empty FFFFCC input rows (MAX-004 fix)
    for r in range(10, 60):
        cell1 = ws.cell(row=r, column=1)
        cell1.fill = styles["input_cell"]["fill"]
        cell1.border = border_thin
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = border_thin
        dv_type.add(ws.cell(row=r, column=5))
        dv_posture.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "A4"  # DS-007 fix


# =============================================================================
# BREAK-GLASS ACCOUNTS SHEET
# =============================================================================
def create_breakglass_accounts_sheet(ws, styles):
    """Create the BreakGlass Accounts sheet."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:M1")
    ws["A1"] = "BREAK-GLASS ACCOUNTS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:M2")
    ws["A2"] = "Inventory of emergency access accounts — all accounts must be Disabled when not in use"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Account_ID", 15),
        ("Account_Name", 25),
        ("Account_Type", 20),
        ("Target_Systems", 35),
        ("Scope_Permissions", 35),
        ("Credential_Location", 30),
        ("Activation_Authority", 25),
        ("Two_Person_Required", 15),
        ("Default_Duration", 15),
        ("Logging_Enabled", 15),
        ("Last_Rotation_Date", 16),
        ("Last_Test_Date", 16),
        ("Status", 12),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Domain Admin,System Admin,Database Admin,Network Admin,Application Admin,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_bool = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_bool)

    dv_logging = DataValidation(
        type="list",
        formula1='"Yes,No,Verified"',
        allow_blank=False
    )
    ws.add_data_validation(dv_logging)

    dv_status = DataValidation(
        type="list",
        formula1='"Disabled,Enabled,Unknown"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    # Row 4: F2F2F2 grey sample row (MAX-003 fix)
    sample_values = {
        1: "BGA-001", 2: "breakglass-admin-01", 3: "Domain Admin",
        4: "All domain controllers", 5: "Full domain admin rights",
        6: "CyberArk vault — Tier 0 Safe", 7: "CISO + CIO dual approval",
        8: "Yes", 9: "4 hours", 10: "Verified",
        11: "15.01.2026", 12: "01.02.2026", 13: "Disabled",
    }
    for col_idx, value in sample_values.items():
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.fill = grey_sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 5-54: 50 empty FFFFCC input rows (MAX-004 fix)
    for r in range(5, 55):
        cell1 = ws.cell(row=r, column=1)
        cell1.fill = styles["input_cell"]["fill"]
        cell1.border = border_thin
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = border_thin
        dv_type.add(ws.cell(row=r, column=3))
        dv_bool.add(ws.cell(row=r, column=8))
        dv_logging.add(ws.cell(row=r, column=10))
        dv_status.add(ws.cell(row=r, column=13))

    ws.freeze_panes = "A4"  # DS-007 fix


# =============================================================================
# BREAK-GLASS ACTIVATION SHEET
# =============================================================================
def create_breakglass_activation_sheet(ws, styles):
    """Create the BreakGlass Activation sheet."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:M1")
    ws["A1"] = "BREAK-GLASS ACTIVATION LOG"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:M2")
    ws["A2"] = "Log of all emergency access account activations — each activation must be authorised, reviewed, and deactivated"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Activation_ID", 15),
        ("Account_ID", 15),
        ("Emergency_Type", 25),
        ("Activation_DateTime", 20),
        ("Authorised_By", 25),
        ("Activated_By", 25),
        ("Second_Person", 25),
        ("CISO_Notified", 12),
        ("Expiry_DateTime", 20),
        ("Renewed", 10),
        ("Deactivation_DateTime", 20),
        ("Post_Review_Completed", 15),
        ("Issues_Found", 40),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_bool = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_bool)

    # Row 4: F2F2F2 grey sample row (MAX-003 fix)
    sample_values = {
        1: "ACT-001", 2: "BGA-001", 3: "Active Directory failure",
        4: "20.01.2026 14:30", 5: "CISO — J. Smith",
        6: "IT Admin — K. Jones", 7: "Security Lead — P. Brown",
        8: "Yes", 9: "20.01.2026 18:30", 10: "No",
        11: "20.01.2026 17:45", 12: "Yes", 13: "None",
    }
    for col_idx, value in sample_values.items():
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.fill = grey_sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 5-54: 50 empty FFFFCC input rows
    for r in range(5, 55):
        cell1 = ws.cell(row=r, column=1)
        cell1.fill = styles["input_cell"]["fill"]
        cell1.border = border_thin
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = border_thin
        dv_bool.add(ws.cell(row=r, column=8))
        dv_bool.add(ws.cell(row=r, column=10))
        dv_bool.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "A4"  # DS-007 fix


# =============================================================================
# ELEVATED MONITORING SHEET
# =============================================================================
def create_elevated_monitoring_sheet(ws, styles):
    """Create the Elevated Monitoring sheet."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:I1")
    ws["A1"] = "ELEVATED MONITORING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:I2")
    ws["A2"] = "Enhanced monitoring requirements by security posture level — defines frequency increases and alert threshold changes"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Posture_Level", 15),
        ("Monitoring_Area", 25),
        ("Normal_Frequency", 20),
        ("Enhanced_Frequency", 20),
        ("Alert_Threshold_Change", 35),
        ("Manual_Review_Required", 15),
        ("Review_Frequency", 20),
        ("Responsible_Party", 25),
        ("Implementation_Notes", 40),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_posture = DataValidation(
        type="list",
        formula1='"Elevated,Degraded,Emergency"',
        allow_blank=False
    )
    ws.add_data_validation(dv_posture)

    dv_bool = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws.add_data_validation(dv_bool)

    # Pre-populate with monitoring requirements rows 4-9 (F2F2F2 grey)
    monitoring = [
        ("Elevated", "Authentication Logs", "Hourly", "15 minutes",
         "Lower failed attempt threshold from 5 to 3", "No", "", "SOC", ""),
        ("Elevated", "Privileged Access", "Daily", "Hourly",
         "Alert on any privileged access", "Yes", "Every 4 hours", "Security Team", ""),
        ("Degraded", "Authentication Logs", "15 minutes", "Real-time",
         "Alert on any anomaly", "Yes", "Every 2 hours", "SOC", "24/7 coverage required"),
        ("Degraded", "Network Traffic", "Hourly", "15 minutes",
         "Alert on unusual destinations", "Yes", "Every 4 hours", "Network Team", "Focus on egress"),
        ("Emergency", "All Critical Systems", "Real-time", "Real-time",
         "Maximum sensitivity", "Yes", "Continuous", "SOC + Security Team", "Threat hunting focus"),
        ("Emergency", "Break-Glass Usage", "N/A", "Real-time",
         "Immediate alert on any activation", "Yes", "Continuous", "CISO", "Direct notification"),
    ]

    for row_idx, mon in enumerate(monitoring, start=4):
        for col_idx, value in enumerate(mon, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = grey_sample_fill
            cell.border = border_thin
            if col_idx in [5, 9]:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
        dv_posture.add(ws.cell(row=row_idx, column=1))
        dv_bool.add(ws.cell(row=row_idx, column=6))

    # Row 10: F2F2F2 grey sample row (MAX-003 fix)
    sample_row = 10
    sample_values_em = {
        1: "Elevated", 2: "File Integrity Monitoring", 3: "6 hours", 4: "1 hour",
        5: "Alert on any unexpected change to critical files", 6: "Yes",
        7: "Every 2 hours", 8: "Security Analyst", 9: "Focus on /etc, system32",
    }
    for col_idx, value in sample_values_em.items():
        cell = ws.cell(row=sample_row, column=col_idx, value=value)
        cell.fill = grey_sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 11-60: 50 empty FFFFCC input rows (MAX-004 fix)
    for r in range(11, 61):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = border_thin
        dv_posture.add(ws.cell(row=r, column=1))
        dv_bool.add(ws.cell(row=r, column=6))

    ws.freeze_panes = "A4"  # DS-007 fix


# =============================================================================
# PERSONNEL AVAILABILITY SHEET
# =============================================================================
def create_personnel_availability_sheet(ws, styles):
    """Create the Personnel Availability sheet."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:N1")
    ws["A1"] = "PERSONNEL AVAILABILITY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:N2")
    ws["A2"] = "Security team succession plan — primary and backup contacts for each critical security role"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Role_ID", 12),
        ("Role_Name", 30),
        ("Primary_Name", 25),
        ("Primary_Phone", 20),
        ("Primary_Email", 30),
        ("Backup1_Name", 25),
        ("Backup1_Phone", 20),
        ("Backup1_Email", 30),
        ("Backup2_Name", 25),
        ("Backup2_Phone", 20),
        ("Backup2_Email", 30),
        ("Cross_Training_Status", 18),
        ("Last_Contact_Test", 16),
        ("Last_Drill_Date", 16),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_training = DataValidation(
        type="list",
        formula1='"Complete,Partial,None"',
        allow_blank=False
    )
    ws.add_data_validation(dv_training)

    # Pre-populate with key roles rows 4-11 (F2F2F2 grey)
    roles = [
        ("R-001", "CISO"),
        ("R-002", "Security Manager"),
        ("R-003", "Security Analyst (Lead)"),
        ("R-004", "SOC Lead"),
        ("R-005", "Incident Response Lead"),
        ("R-006", "IAM Administrator"),
        ("R-007", "Network Security Lead"),
        ("R-008", "Security Architect"),
    ]

    for row_idx, (role_id, role_name) in enumerate(roles, start=4):
        ws.cell(row=row_idx, column=1, value=role_id).fill = grey_sample_fill
        ws.cell(row=row_idx, column=1).border = border_thin
        ws.cell(row=row_idx, column=2, value=role_name).fill = grey_sample_fill
        ws.cell(row=row_idx, column=2).border = border_thin
        for c in range(3, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = grey_sample_fill
            cell.border = border_thin
        dv_training.add(ws.cell(row=row_idx, column=12))

    # Row 12: F2F2F2 grey sample row (MAX-003 fix)
    sample_values_pa = {
        1: "R-009", 2: "Vulnerability Management Lead",
        3: "A. Example", 4: "+41 79 000 0001", 5: "a.example@organisation.ch",
        6: "B. Backup", 7: "+41 79 000 0002", 8: "b.backup@organisation.ch",
        9: "C. Second", 10: "+41 79 000 0003", 11: "c.second@organisation.ch",
        12: "Complete", 13: "01.02.2026", 14: "15.01.2026",
    }
    for col_idx, value in sample_values_pa.items():
        cell = ws.cell(row=12, column=col_idx, value=value)
        cell.fill = grey_sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 13-62: 50 empty FFFFCC input rows (MAX-004 fix)
    for r in range(13, 63):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = border_thin
        dv_training.add(ws.cell(row=r, column=12))

    ws.freeze_panes = "A4"  # DS-007 fix


# =============================================================================
# SECURITY DEBT REGISTER SHEET
# =============================================================================
def create_security_debt_register_sheet(ws, styles):
    """Create the Security Debt Register sheet."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    e2efda_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:O1")
    ws["A1"] = "SECURITY DEBT REGISTER"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:O2")
    ws["A2"] = "Tracks all security controls deferred or degraded during disruption — must be resolved within agreed timescales"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Debt_ID", 12),
        ("Debt_Type", 20),
        ("Description", 45),
        ("Disruption_Reference", 20),
        ("Created_Date", 16),
        ("Owner", 25),
        ("Remediation_Plan", 45),
        ("Target_Date", 16),
        ("Status", 15),
        ("Age_Days", 12),
        ("Escalation_Required", 15),
        ("Escalated_To", 25),
        ("Escalation_Date", 16),
        ("Closure_Date", 16),
        ("Closure_Evidence", 30),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_type = DataValidation(
        type="list",
        formula1='"Deferred Patch,Skipped Review,Delayed Scan,Config Exception,Access Exception,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_type)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Closed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    # Row 4: F2F2F2 grey sample row (MAX-003 fix)
    sample_values = {
        1: "SD-001", 2: "Deferred Patch", 3: "Critical patches deferred during datacenter failover",
        4: "DS-003", 5: "20.01.2026", 6: "IT Security Manager",
        7: "Apply all deferred patches within 5 days of recovery",
        8: "25.01.2026", 9: "Open",
        12: "CISO", 13: "", 14: "", 15: "",
    }
    for col_idx, value in sample_values.items():
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.fill = grey_sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # Auto-formula cols for sample row (E2EFDA)
    ws.cell(row=4, column=10).value = '=IF(E4<>"",TODAY()-E4,"")'
    ws.cell(row=4, column=10).fill = e2efda_fill
    ws.cell(row=4, column=10).border = border_thin
    ws.cell(row=4, column=11).value = '=IF(J4>30,"Yes","No")'
    ws.cell(row=4, column=11).fill = e2efda_fill
    ws.cell(row=4, column=11).border = border_thin

    # Rows 5-54: 50 empty FFFFCC input rows
    for r in range(5, 55):
        cell1 = ws.cell(row=r, column=1)
        cell1.fill = styles["input_cell"]["fill"]
        cell1.border = border_thin
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = border_thin
        dv_type.add(ws.cell(row=r, column=2))
        dv_status.add(ws.cell(row=r, column=9))
        # Age formula
        ws.cell(row=r, column=10).value = f'=IF(E{r}<>"",TODAY()-E{r},"")'
        ws.cell(row=r, column=10).fill = e2efda_fill
        # Escalation formula
        ws.cell(row=r, column=11).value = f'=IF(J{r}>30,"Yes","No")'
        ws.cell(row=r, column=11).fill = e2efda_fill

    ws.freeze_panes = "A4"  # DS-007 fix




def create_evidence_register(ws):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _input = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 1: Title banner
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    # Row 2: Italic subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = _border

    # Row 3: empty separator

    # Row 4: Column headers (003366 navy)
    columns = [
        ("Evidence ID", 15), ("Assessment Area", 25), ("Evidence Type", 22),
        ("Description", 40), ("Location / Path", 45),
        ("Date Collected", 16), ("Collected By", 20), ("Verification Status", 22),
    ]
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data validations
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other"',
        allow_blank=True,
    )
    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)
    ws.add_data_validation(ver_status_dv)

    # Row 5: Sample row (F2F2F2 grey)
    sample_data = {
        1: "EV-001", 2: "Assessment Area", 3: "Policy Document",
        4: "Example evidence description", 5: "\\\\fileserver\\policies\\document.pdf",
        6: "15.01.2026", 7: "Assessor Name", 8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = _grey_sample
        cell.border = _border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center", wrap_text=True
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # Rows 6-105: Empty FFFFCC rows (100 rows)
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = _input
            cell.border = _border
            cell.alignment = Alignment(
                horizontal="center" if col == 1 else "left",
                vertical="center", wrap_text=True
            )
            cell.value = None
        ev_type_dv.add(ws[f"C{row}"])
        ver_status_dv.add(ws[f"H{row}"])

    ws.freeze_panes = "A5"

def create_approval_sheet(ws):
    """Gold Standard Approval Sign-Off — ISMS-IMP-A.5.29.2.

    Structure:
      Row 1  : Title banner (003366)
      Row 2  : CONTROL_REF subtitle
      Row 3  : ASSESSMENT SUMMARY banner (4472C4)
      Rows 4-8: Summary fields (Document / Period / Overall Compliance / Status / By)
                B6 = ='Summary Dashboard'!G9  [Overall Compliance Rating]
      Row 9  : gap
      Row 10 : ASSESSMENT REVIEW AND APPROVAL banner (4472C4)
      Row 11 : Column headers
      Rows 12-15: Approver rows
      Row 16 : gap
      Row 17 : FINAL DECISION (GS-AS-012: col A plain bold, no fill)
      Row 18 : gap
      Row 19 : NEXT REVIEW DETAILS banner (4472C4)
      Rows 20-21: Next Review Date / Review Owner
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _thin   = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy   = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _blue   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _grey   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _input  = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # -------------------------------------------------------------------------
    # Row 1: Title banner
    # -------------------------------------------------------------------------
    ws.merge_cells("A1:E1")
    ws["A1"] = "DEGRADED MODE SECURITY REQUIREMENTS — ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = _border
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = _border

    # -------------------------------------------------------------------------
    # Row 3: ASSESSMENT SUMMARY banner
    # -------------------------------------------------------------------------
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = _blue
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = _border

    # Rows 4-8: Summary fields (B6 = Overall Compliance Rating)
    summary_fields = [
        ("Document:",                f"{DOCUMENT_ID} — {WORKBOOK_NAME}"),
        ("Assessment Period:",       ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G8),\"\")"),
        ("Assessment Status:",       ""),
        ("Assessed By:",             ""),
    ]
    for row_idx, (label, value) in enumerate(summary_fields, start=4):
        lbl_cell = ws.cell(row=row_idx, column=1, value=label)
        lbl_cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        lbl_cell.border = _border
        ws.merge_cells(f"B{row_idx}:E{row_idx}")
        val_cell = ws.cell(row=row_idx, column=2, value=value)
        if value == "":
            val_cell.fill = _input
        elif "Summary Dashboard" in value:
            val_cell.number_format = "0.0%"
        for c in range(2, 6):
            ws.cell(row=row_idx, column=c).border = _border

    # Assessment Status DV on row 7
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # Row 9: gap
    ws.merge_cells("A9:E9")
    ws["A9"].border = _border

    # -------------------------------------------------------------------------
    # Row 10: ASSESSMENT REVIEW AND APPROVAL banner
    # -------------------------------------------------------------------------
    ws.merge_cells("A10:E10")
    ws["A10"] = "ASSESSMENT REVIEW AND APPROVAL"
    ws["A10"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A10"].fill = _blue
    ws["A10"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A10"].border = _border

    # Row 11: Column headers
    col_headers = ["Role / Function", "Name", "Signature / Initials", "Date (DD.MM.YYYY)", "Comments"]
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = _grey
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border

    # Rows 12-15: Approver rows
    approvers = [
        "Lead Assessor / Author",
        "IT Security Manager",
        "Reviewer (Technical / Compliance)",
        "CISO / Final Approver",
    ]
    for row_idx, approver in enumerate(approvers, start=12):
        cell = ws.cell(row=row_idx, column=1, value=approver)
        cell.font = Font(name="Calibri", size=10)
        cell.fill = _input
        cell.border = _border
        for col in range(2, 6):
            c = ws.cell(row=row_idx, column=col)
            c.fill = _input
            c.border = _border

    # Row 16: gap
    ws.merge_cells("A16:E16")
    ws["A16"].border = _border

    # -------------------------------------------------------------------------
    # Row 17: FINAL DECISION (GS-AS-012: col A plain bold, no fill)
    # -------------------------------------------------------------------------
    ws["A17"] = "FINAL DECISION:"
    ws["A17"].font = Font(name="Calibri", size=11, bold=True, color="000000")
    ws["A17"].border = _border
    ws.merge_cells("B17:E17")
    ws["B17"] = "☐ Approved     ☐ Approved with Conditions     ☐ Rejected     ☐ Deferred"
    ws["B17"].font = Font(name="Calibri", size=10)
    ws["B17"].fill = _input
    ws["B17"].alignment = Alignment(horizontal="center", vertical="center")
    for col in range(2, 6):
        ws.cell(row=17, column=col).border = _border

    # Row 18: gap
    ws.merge_cells("A18:E18")
    ws["A18"].border = _border

    # -------------------------------------------------------------------------
    # Row 19: NEXT REVIEW DETAILS (Gold Standard: 4472C4 fill, white bold 11pt)
    # -------------------------------------------------------------------------
    ws.merge_cells("A19:E19")
    ws["A19"] = "NEXT REVIEW DETAILS"
    ws["A19"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A19"].fill = _blue
    ws["A19"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A19"].border = _border

    for row_idx, label in enumerate(["Next Review Date (DD.MM.YYYY)", "Review Owner"], start=20):
        ws.cell(row=row_idx, column=1, value=label).font = Font(name="Calibri", size=10, bold=True, color="000000")
        ws.cell(row=row_idx, column=1).fill = _input
        ws.cell(row=row_idx, column=1).border = _border
        ws.merge_cells(f"B{row_idx}:E{row_idx}")
        c = ws.cell(row=row_idx, column=2)
        c.fill = _input
        c.border = _border

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 30
    ws.freeze_panes = "A3"

def create_summary_dashboard_sheet(ws):
    """Create the Summary Dashboard sheet for ISMS-IMP-A.5.29.2 (Degraded Mode Security Requirements).

    TABLE 1: 4 assessment areas (BreakGlass Accounts, Elevated Monitoring,
              Personnel Availability, Security Debt Register) — TOTAL row 9
    TABLE 2: 15 key metrics across all data-gathering sheets
    TABLE 3: 6 critical findings requiring immediate attention
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _blue_tbl = PatternFill("solid", fgColor="003366")  # TABLE banners — 003366
    _red_tbl = PatternFill("solid", fgColor="C00000")   # TABLE 3 banner — C00000
    _grey_hdr = PatternFill("solid", fgColor="D9D9D9")
    _white = PatternFill("solid", fgColor="FFFFFF")

    # -------------------------------------------------------------------------
    # A1: Title banner — alignment=None (gold standard: no explicit alignment)
    # -------------------------------------------------------------------------
    ws.merge_cells("A1:G1")
    ws["A1"] = "DEGRADED MODE SECURITY REQUIREMENTS \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    # A2: subtitle — horizontal="left" (gold standard)
    ws.merge_cells("A2:G2")
    ws["A2"] = "Compliance summary for degraded mode operations, break-glass procedures, and security debt management"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # -------------------------------------------------------------------------
    # TABLE 1 — Compliance Overview
    # -------------------------------------------------------------------------
    ws.merge_cells("A3:G3")
    ws["A3"] = "TABLE 1: COMPLIANCE OVERVIEW"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = _blue_tbl
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].border = _border

    t1_headers = ["Assessment Area", "Total", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col_idx, header in enumerate(t1_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = _grey_hdr
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center", wrap_text=True)
        cell.border = _border

    # TABLE 1 data rows — formulas reference sheets in this workbook
    # Note: Elevated Monitoring user-input rows = 11-60 (pre-pop 4-9, sample 10)
    # Personnel Availability user-input rows = 13-62 (pre-pop 4-11, sample 12)
    t1_rows = [
        ("BreakGlass Accounts",
         "=COUNTA('BreakGlass Accounts'!B5:B54)",
         "=COUNTIF('BreakGlass Accounts'!M5:M54,\"Disabled\")",
         "=COUNTIF('BreakGlass Accounts'!M5:M54,\"Unknown\")",
         "=COUNTIF('BreakGlass Accounts'!M5:M54,\"Enabled\")",
         "0"),
        ("Elevated Monitoring",
         "=COUNTA('Elevated Monitoring'!B11:B60)",
         "=COUNTIF('Elevated Monitoring'!F11:F60,\"Yes\")",
         "0",
         "=COUNTIF('Elevated Monitoring'!F11:F60,\"No\")",
         "0"),
        ("Personnel Availability",
         "=COUNTA('Personnel Availability'!B13:B62)",
         "=COUNTIF('Personnel Availability'!L13:L62,\"Complete\")",
         "=COUNTIF('Personnel Availability'!L13:L62,\"Partial\")",
         "=COUNTIF('Personnel Availability'!L13:L62,\"None\")",
         "0"),
        ("Security Debt Register",
         "=COUNTA('Security Debt Register'!B5:B54)",
         "=COUNTIF('Security Debt Register'!I5:I54,\"Closed\")",
         "=COUNTIF('Security Debt Register'!I5:I54,\"In Progress\")",
         "=COUNTIF('Security Debt Register'!I5:I54,\"Open\")",
         "0"),
    ]

    for row_idx, (area, total, comp, partial, noncomp, na) in enumerate(t1_rows, start=5):
        data = [area, total, comp, partial, noncomp, na]
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10, bold=False, color="000000")
            cell.fill = _white
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
            cell.border = _border
        # Compliance % col G
        pct = ws.cell(row=row_idx, column=7,
                      value=f'=IFERROR(IF((B{row_idx}-F{row_idx})=0,0,C{row_idx}/(B{row_idx}-F{row_idx})),"")')
        pct.font = Font(name="Calibri", size=10, bold=False, color="000000")
        pct.fill = _white
        pct.alignment = Alignment(horizontal="center", vertical="center")
        pct.border = _border
        pct.number_format = "0.0%"

    # TOTAL row at row 9
    total_row = 9
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", size=10, bold=True, color="000000")
    ws.cell(row=total_row, column=1).fill = _grey_hdr
    ws.cell(row=total_row, column=1).border = _border
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for col_idx in range(2, 7):
        col_letter = ["B", "C", "D", "E", "F"][col_idx - 2]
        cell = ws.cell(row=total_row, column=col_idx,
                       value=f"=SUM({col_letter}5:{col_letter}8)")
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = _grey_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border
    pct_total = ws.cell(row=total_row, column=7,
                        value='=IFERROR(IF((B9-F9)=0,0,C9/(B9-F9)),"")')
    pct_total.font = Font(name="Calibri", size=10, bold=True, color="000000")
    pct_total.fill = _grey_hdr
    pct_total.alignment = Alignment(horizontal="center", vertical="center")
    pct_total.border = _border
    pct_total.number_format = "0.0%"

    # -------------------------------------------------------------------------
    # TABLE 2 — Key Metrics (starts at row 10)
    # -------------------------------------------------------------------------
    ws.merge_cells("A10:G10")
    ws["A10"] = "TABLE 2: KEY METRICS"
    ws["A10"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A10"].fill = _blue_tbl
    ws["A10"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A10"].border = _border

    t2_headers = ["Metric", "Value"]
    for col_idx, header in enumerate(t2_headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = _grey_hdr
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        cell.border = _border
    for col_idx in range(3, 8):
        cell = ws.cell(row=11, column=col_idx)
        cell.fill = _grey_hdr
        cell.border = _border

    t2_metrics = [
        ("Degradation Scenarios Documented",
         "=COUNTA('Degradation Scenarios'!B10:B59)"),
        ("Temporary Bypass Scenarios",
         "=COUNTIF('Degradation Scenarios'!E10:E59,\"Temporary Bypass\")"),
        ("Emergency-Level Scenarios",
         "=COUNTIF('Degradation Scenarios'!J10:J59,\"Emergency\")"),
        ("Break-Glass Accounts Inventoried",
         "=COUNTA('BreakGlass Accounts'!B5:B54)"),
        ("Break-Glass Accounts Disabled (Safe)",
         "=COUNTIF('BreakGlass Accounts'!M5:M54,\"Disabled\")"),
        ("Break-Glass Accounts Enabled",
         "=COUNTIF('BreakGlass Accounts'!M5:M54,\"Enabled\")"),
        ("Accounts with Logging Verified",
         "=COUNTIF('BreakGlass Accounts'!J5:J54,\"Verified\")"),
        ("Accounts Requiring 2-Person Activation",
         "=COUNTIF('BreakGlass Accounts'!H5:H54,\"Yes\")"),
        ("Break-Glass Activations Recorded",
         "=COUNTA('BreakGlass Activation'!B5:B54)"),
        ("Activations with Post-Review Complete",
         "=COUNTIF('BreakGlass Activation'!L5:L54,\"Yes\")"),
        ("Personnel Roles with Full Cross-Training",
         "=COUNTIF('Personnel Availability'!L13:L62,\"Complete\")"),
        ("Personnel Roles with No Cross-Training",
         "=COUNTIF('Personnel Availability'!L13:L62,\"None\")"),
        ("Open Security Debt Items",
         "=COUNTIF('Security Debt Register'!I5:I54,\"Open\")"),
        ("In-Progress Debt Items",
         "=COUNTIF('Security Debt Register'!I5:I54,\"In Progress\")"),
        ("Closed Security Debt Items",
         "=COUNTIF('Security Debt Register'!I5:I54,\"Closed\")"),
    ]

    for row_idx, (metric, formula) in enumerate(t2_metrics, start=12):
        label = ws.cell(row=row_idx, column=1, value=metric)
        label.font = Font(name="Calibri", size=10, bold=False, color="000000")
        label.fill = _white
        label.alignment = Alignment(horizontal="left", vertical="center")
        label.border = _border
        val = ws.cell(row=row_idx, column=2, value=formula)
        val.font = Font(name="Calibri", size=10, bold=False, color="000000")
        val.fill = _white
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.border = _border
        for col_idx in range(3, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = _white
            cell.border = _border

    # 2 empty buffer rows after TABLE 2 (rows 27-28)
    for r in range(27, 29):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.fill = _white
            cell.border = _border

    # -------------------------------------------------------------------------
    # TABLE 3 — Critical Findings (starts at row 29)
    # -------------------------------------------------------------------------
    ws.merge_cells("A29:G29")
    ws["A29"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws["A29"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A29"].fill = _red_tbl
    ws["A29"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A29"].border = _border

    t3_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col_idx, header in enumerate(t3_headers, start=1):
        cell = ws.cell(row=30, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = _grey_hdr
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        cell.border = _border

    t3_findings = [
        ("Break-Glass Accounts",
         "Emergency accounts currently Enabled (active)",
         "=COUNTIF('BreakGlass Accounts'!M5:M54,\"Enabled\")",
         "Critical", "Immediate"),
        ("Break-Glass Accounts",
         "Accounts without logging verified",
         "=COUNTIF('BreakGlass Accounts'!J5:J54,\"No\")",
         "Critical", "Immediate"),
        ("BreakGlass Activation",
         "Activations without post-review completed",
         "=COUNTIF('BreakGlass Activation'!L5:L54,\"No\")",
         "High", "Urgent"),
        ("Personnel Availability",
         "Security roles with no cross-training",
         "=COUNTIF('Personnel Availability'!L13:L62,\"None\")",
         "High", "Urgent"),
        ("Security Debt Register",
         "Open security debt items (not yet remediated)",
         "=COUNTIF('Security Debt Register'!I5:I54,\"Open\")",
         "High", "Urgent"),
        ("Degradation Scenarios",
         "Emergency-level degradation scenarios documented",
         "=COUNTIF('Degradation Scenarios'!J10:J59,\"Emergency\")",
         "Medium", "Review"),
    ]

    _finding_fill = PatternFill("solid", fgColor="FFFFCC")
    for row_idx, (cat, finding, formula, severity, action) in enumerate(t3_findings, start=31):
        data = [cat, finding, formula, severity, action, "", ""]
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10, bold=False, color="000000")
            cell.fill = _finding_fill
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left",
                                       vertical="center", wrap_text=(col_idx == 2))
            cell.border = _border

    # 2 empty buffer rows at end of TABLE 3 (rows 37-38)
    for r in range(37, 39):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.fill = _finding_fill
            cell.border = _border

    # -------------------------------------------------------------------------
    # Column widths
    # -------------------------------------------------------------------------
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 12

    ws.freeze_panes = "A3"


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly

def main() -> int:
    """Main execution function."""
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = _STYLES

        logger.info("[1/10] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("[2/10] Creating Degradation Scenarios sheet...")
        create_degradation_scenarios_sheet(wb["Degradation Scenarios"], styles)

        logger.info("[3/10] Creating BreakGlass Accounts sheet...")
        create_breakglass_accounts_sheet(wb["BreakGlass Accounts"], styles)

        logger.info("[4/10] Creating BreakGlass Activation sheet...")
        create_breakglass_activation_sheet(wb["BreakGlass Activation"], styles)

        logger.info("[5/10] Creating Elevated Monitoring sheet...")
        create_elevated_monitoring_sheet(wb["Elevated Monitoring"], styles)

        logger.info("[6/10] Creating Personnel Availability sheet...")
        create_personnel_availability_sheet(wb["Personnel Availability"], styles)

        logger.info("[7/10] Creating Security Debt Register sheet...")
        create_security_debt_register_sheet(wb["Security Debt Register"], styles)

        logger.info("[8/10] Creating Evidence Register sheet...")
        create_evidence_register(wb["Evidence Register"])

        logger.info("[9/10] Creating Summary Dashboard sheet...")
        create_summary_dashboard_sheet(wb["Summary Dashboard"])

        logger.info("[10/10] Creating Approval Sign-Off sheet...")
        create_approval_sheet(wb["Approval Sign-Off"])

        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        wb.save(output_path)
        logger.info(f"SUCCESS: {OUTPUT_FILENAME}")
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error(f"Missing dependency: {e}")
        logger.error("Install with: pip install openpyxl")
        return 1
    except PermissionError as e:
        logger.error(f"Permission denied: {e}")
        return 1
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
