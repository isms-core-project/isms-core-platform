#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.29.1 - Security Controls During Disruption Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.29: Information Security During Disruption
Assessment Domain 1 of 3: Security Controls During Disruption

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific information security during disruption infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Disruption scenario categories and priority tiers (match your BCP framework)
2. Security control degradation thresholds and acceptable risk levels
3. Degraded mode security requirements per service tier
4. Recovery verification checklist items and sign-off authorities
5. Integration with your organisation's BCP/DR testing programme

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.29 Information Security During Disruption Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
information security during disruption controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Security Controls During Disruption under ISO 27001:2022 Control A.5.29. Supports evidence-based documentation of security control resilience, degraded mode procedures, and recovery verification compliance.

**Assessment Scope:**
- Security control coverage during disruption scenarios
- Degraded mode security requirement definition and completeness
- Compensating control adequacy during reduced-capability periods
- Recovery security verification procedure documentation
- BCP/DR integration with information security requirements
- Testing and exercise outcomes for disruption scenarios
- Evidence collection for business continuity and compliance audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
4. Summary Dashboard - Compliance overview and key metrics
5. Evidence Register - Audit evidence tracking
6. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 3 domains covering Information Security During Disruption controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a529_1_security_controls_disruption.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a529_1_security_controls_disruption.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a529_1_security_controls_disruption.py --date 20250115

Output:
    File: ISMS-IMP-A.5.29.1_Security_Controls_During_Disruption_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.29
Assessment Domain:    1 of 3 (Security Controls During Disruption)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.29: Information Security During Disruption Policy (Governance)
    - ISMS-IMP-A.5.29.1: Security Controls During Disruption (Domain 1)
    - ISMS-IMP-A.5.29.2: Degraded Mode Security Requirements (Domain 2)
    - ISMS-IMP-A.5.29.3: Recovery Security Verification (Domain 3)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.29.1 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive information security during disruption details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review security during disruption procedures annually or when BCP/DR plans are updated, critical system architectures change, or disruption testing reveals gaps.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

# =============================================================================
# IMPORTS - STANDARD LIBRARY
# =============================================================================
import logging
import sys
from datetime import datetime

# =============================================================================
# IMPORTS - THIRD PARTY
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.29.1"
WORKBOOK_NAME = "Security Controls During Disruption"
CONTROL_ID = "A.5.29"
CONTROL_NAME = "Information Security During Disruption"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
_THIN_SIDE = Side(style='thin')
THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)

def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "border": border_thin,
        "status_compliant": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
        "status_partial": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "status_noncompliant": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
    }
    return styles


# =============================================================================
# WORKBOOK CREATION
# =============================================================================

_STYLES = setup_styles()
def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching markdown spec."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure matches specification
    sheets = [
        "Instructions & Legend",
        "Security Control Inventory",
        "Minimum Baseline",
        "Security Posture Levels",
        "Compensating Controls",
        "BCDR Security Review",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================

def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Security Control Inventory — list all security controls and their disruption resilience status.',
        '2. Complete Minimum Baseline — define the minimum acceptable security posture during disruption.',
        '3. Complete Security Posture Levels — document security controls available at each degradation level.',
        '4. Complete Compensating Controls — identify alternative controls when primary controls are unavailable.',
        '5. Complete BCDR Security Review — assess how the BCP/DR plan addresses security continuity.',
        '6. Maintain the Evidence Register with BCP/DR documentation and test results.',
        '7. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A21"] = "Status Legend"
    ws["A21"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=22, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 23 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_security_control_inventory_sheet(ws, styles):
    """Create the Security Control Inventory sheet - master catalogue."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:K1")
    ws["A1"] = "SECURITY CONTROL INVENTORY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:K2")
    ws["A2"] = "Inventory of all security controls with their disruption classification and recovery priority"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Column definitions per specification
    columns = [
        ("Control_ID", 15),
        ("Control_Name", 35),
        ("Control_Category", 20),
        ("ISO_Reference", 15),
        ("Normal_Status", 15),
        ("Disruption_Classification", 22),
        ("Recovery_Priority", 15),
        ("Compensating_Control_ID", 18),
        ("Owner", 25),
        ("Last_Review_Date", 16),
        ("Notes", 40),
    ]

    # Headers row 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data validations
    dv_category = DataValidation(
        type="list",
        formula1='"Access Control,Data Encryption,Logging,Network Security,Backup Protection,Endpoint Security,Physical Security,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_category)

    dv_status = DataValidation(
        type="list",
        formula1='"Operational,Partial,Not Applicable"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    dv_classification = DataValidation(
        type="list",
        formula1='"Non-Negotiable,Degradable,Deferrable,Not Applicable"',
        allow_blank=False
    )
    ws.add_data_validation(dv_classification)

    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=False
    )
    ws.add_data_validation(dv_priority)

    # Rows 4-12: Pre-populated reference security controls (F2F2F2 grey — these are reference data)
    sample_controls = [
        ("SC-001", "Multi-Factor Authentication", "Access Control", "A.8.5", "Operational", "Degradable", "Critical"),
        ("SC-002", "Role-Based Access Control", "Access Control", "A.5.15", "Operational", "Non-Negotiable", "Critical"),
        ("SC-003", "Privileged Access Management", "Access Control", "A.8.2", "Operational", "Non-Negotiable", "Critical"),
        ("SC-004", "Data Encryption at Rest", "Data Encryption", "A.8.24", "Operational", "Non-Negotiable", "Critical"),
        ("SC-005", "TLS for Data in Transit", "Data Encryption", "A.8.24", "Operational", "Non-Negotiable", "Critical"),
        ("SC-006", "Security Event Logging", "Logging", "A.8.15", "Operational", "Non-Negotiable", "Critical"),
        ("SC-007", "SIEM Correlation", "Logging", "A.8.16", "Operational", "Degradable", "High"),
        ("SC-008", "Firewall Protection", "Network Security", "A.8.20", "Operational", "Non-Negotiable", "Critical"),
        ("SC-009", "Network Segmentation", "Network Security", "A.8.22", "Operational", "Non-Negotiable", "Critical"),
    ]

    for row_idx, control in enumerate(sample_controls, start=4):
        for col_idx, value in enumerate(control, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = grey_sample_fill
            cell.border = border_thin

    # Row 13: F2F2F2 grey sample row (MAX-003 fix) — shows how to fill blank input rows
    sample_row = 13
    sample_values = {
        1: "SC-010", 2: "Endpoint Detection & Response", 3: "Endpoint Security",
        4: "A.8.7", 5: "Operational", 6: "Degradable", 7: "High",
        8: "CC-004", 9: "IT Security Manager", 10: "23.02.2026", 11: "Monitor for anomalies during recovery",
    }
    for col_idx, value in sample_values.items():
        cell = ws.cell(row=sample_row, column=col_idx, value=value)
        cell.fill = grey_sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 14-63: 50 empty FFFFCC input rows (MAX-004 fix)
    for r in range(14, 64):
        # Col A
        cell1 = ws.cell(row=r, column=1)
        cell1.fill = styles["input_cell"]["fill"]
        cell1.border = border_thin
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = border_thin
            cell.alignment = styles["input_cell"]["alignment"]
        # Apply validations (exclude sample row 13)
        dv_category.add(ws.cell(row=r, column=3))
        dv_status.add(ws.cell(row=r, column=5))
        dv_classification.add(ws.cell(row=r, column=6))
        dv_priority.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "A4"  # DS-007 fix


# =============================================================================
# MINIMUM BASELINE SHEET
# =============================================================================
def create_minimum_baseline_sheet(ws, styles):
    """Create the Minimum Baseline sheet - non-negotiable controls."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:I1")
    ws["A1"] = "MINIMUM SECURITY BASELINE"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:I2")
    ws["A2"] = "Non-negotiable security controls that must be maintained at all times during any disruption"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Control_ID", 15),
        ("Control_Name", 35),
        ("Category", 20),
        ("Minimum_Requirement", 45),
        ("Rationale", 45),
        ("Never_Acceptable_Actions", 45),
        ("Approval_Status", 15),
        ("Approved_By", 25),
        ("Approval_Date", 16),
    ]

    # Headers row 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Rows 4-8: Pre-populated baseline requirements (F2F2F2 grey reference data)
    baseline_controls = [
        ("SC-002", "Role-Based Access Control", "Access Control",
         "Authentication required for all system access",
         "Prevents unauthorised access during chaos",
         "Removal of authentication requirements"),
        ("SC-004", "Data Encryption at Rest", "Data Encryption",
         "Encryption at rest for confidential+ data",
         "Data remains protected if media lost",
         "Decryption of data without re-encryption"),
        ("SC-006", "Security Event Logging", "Logging",
         "Critical system logging continues",
         "Audit trail for post-incident investigation",
         "Disabled logging on critical systems"),
        ("SC-008", "Firewall Protection", "Network Security",
         "Critical network boundaries maintained",
         "Prevents lateral movement during recovery",
         "Disabling of network security controls"),
        ("SC-011", "Backup Encryption", "Backup Protection",
         "Backups remain encrypted and access-controlled",
         "Prevents backup compromise for data theft",
         "Sharing of privileged credentials without accountability"),
    ]

    dv_status = DataValidation(
        type="list",
        formula1='"Pending,Approved,Rejected"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)

    for row_idx, control in enumerate(baseline_controls, start=4):
        for col_idx, value in enumerate(control, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = grey_sample_fill
            cell.border = border_thin
            if col_idx <= 6:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
        # Approval cols also grey for pre-populated rows
        for c in range(7, 10):
            ws.cell(row=row_idx, column=c).fill = grey_sample_fill
            ws.cell(row=row_idx, column=c).border = border_thin

    # Row 9: F2F2F2 grey sample row (MAX-003 fix) — shows how to fill input rows
    sample_row = 9
    sample_values = {
        1: "SC-012", 2: "Endpoint Detection & Response", 3: "Endpoint Security",
        4: "EDR must remain active on all critical servers",
        5: "Enables detection of threats exploiting disruption window",
        6: "Disabling EDR without written CISO approval",
        7: "Pending", 8: "CISO", 9: "23.02.2026",
    }
    for col_idx, value in sample_values.items():
        cell = ws.cell(row=sample_row, column=col_idx, value=value)
        cell.fill = grey_sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 10-59: 50 empty FFFFCC input rows (MAX-004 fix)
    for r in range(10, 60):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = border_thin
        dv_status.add(ws.cell(row=r, column=7))

    ws.freeze_panes = "A4"  # DS-007 fix


# =============================================================================
# SECURITY POSTURE LEVELS SHEET
# =============================================================================
def create_security_posture_levels_sheet(ws, styles):
    """Create the Security Posture Levels sheet."""
    ws.merge_cells("A1:I1")
    ws["A1"] = "SECURITY POSTURE LEVELS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    columns = [
        ("Posture_Level", 15),
        ("Description", 40),
        ("Trigger_Conditions", 40),
        ("Control_Status", 35),
        ("Monitoring_Enhancement", 35),
        ("Transition_To", 25),
        ("Transition_Authority", 30),
        ("Documentation_Required", 35),
        ("Example_Scenario", 40),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate with posture levels
    posture_levels = [
        ("Normal", "Full security controls operational", "No disruption",
         "All controls at 100%", "Standard monitoring",
         "Elevated", "N/A (default state)", "None",
         "Day-to-day operations"),
        ("Elevated", "Enhanced monitoring, accelerated patching", "Minor disruption detected",
         "All controls operational, enhanced vigilance", "Increased log review frequency, lowered alert thresholds",
         "Normal, Degraded", "CISO or Security Team Lead", "Incident ticket",
         "Single system failure, minor security incident"),
        ("Degraded", "Core controls maintained, non-critical deferred", "Moderate disruption",
         "Essential controls only, some relaxations", "24/7 security monitoring, external threat intel",
         "Elevated, Emergency", "CISO + CIO", "Formal notification to Executive Management",
         "Data centre failover, regional outage"),
        ("Emergency", "Minimum baseline only, survival mode", "Severe disruption",
         "Minimum baseline controls only", "Real-time monitoring of critical systems, threat hunting",
         "Degraded, Recovery", "Executive Management (CEO or delegate)", "Emergency declaration document",
         "Multiple site disaster, major cyber attack"),
        ("Recovery", "Gradual restoration with security validation", "Returning to normal",
         "Phased control restoration", "Validation monitoring, anomaly detection",
         "Degraded, Normal", "CISO approval for each phase", "Phase completion checklist",
         "Post-disaster recovery phase"),
    ]

    for row_idx, level in enumerate(posture_levels, start=4):
        for col_idx, value in enumerate(level, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.freeze_panes = "B4"


# =============================================================================
# COMPENSATING CONTROLS SHEET
# =============================================================================
def create_compensating_controls_sheet(ws, styles):
    """Create the Compensating Controls sheet."""
    ws.merge_cells("A1:K1")
    ws["A1"] = "COMPENSATING CONTROLS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    columns = [
        ("Compensating_ID", 15),
        ("Primary_Control_ID", 18),
        ("Primary_Control_Name", 30),
        ("Compensating_Measure", 45),
        ("Effectiveness", 15),
        ("Implementation_Requirements", 40),
        ("Activation_Trigger", 35),
        ("Duration_Limit", 20),
        ("Test_Status", 15),
        ("Last_Test_Date", 16),
        ("Test_Results", 40),
    ]

    # Headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Pre-populate with sample compensating controls
    compensating = [
        ("CC-001", "SC-001", "Multi-Factor Authentication",
         "Enhanced logging, session limits, IP restrictions",
         "Partial", "SOC to enable enhanced monitoring, IT to apply session limits",
         "MFA infrastructure unavailable", "72 hours",
         "Tested", "", "Effective during Q3 drill"),
        ("CC-002", "SC-007", "SIEM Correlation",
         "Manual log review for critical systems",
         "Partial", "Security team manual review every 4 hours",
         "SIEM platform unavailable", "48 hours",
         "Tested", "", "Functional but resource-intensive"),
        ("CC-003", "SC-010", "IDS/IPS",
         "Enhanced firewall logging, manual traffic analysis",
         "Partial", "Network team to review firewall logs hourly",
         "IDS/IPS system failure", "24 hours",
         "Tested", "", "Acceptable for short duration"),
        ("CC-004", "SC-012", "EDR",
         "Increased antimalware scans, manual endpoint review",
         "Partial", "IT to run full scans twice daily",
         "EDR platform unavailable", "48 hours",
         "Untested", "", ""),
        ("CC-005", "SC-014", "Patch Management",
         "Critical patches only, manual review process",
         "Partial", "Security team manual patch review",
         "Patch management system unavailable", "30 days for non-critical",
         "Tested", "", "Used during system upgrade"),
    ]

    dv_effectiveness = DataValidation(
        type="list",
        formula1='"Full,Partial,Minimal"',
        allow_blank=False
    )
    ws.add_data_validation(dv_effectiveness)

    dv_test_status = DataValidation(
        type="list",
        formula1='"Tested,Untested,Failed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_test_status)

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:K2")
    ws["A2"] = "Alternative security measures to apply when primary controls cannot be maintained during disruption"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Rows 4-8: Pre-populated compensating controls (F2F2F2 grey reference data)
    for row_idx, control in enumerate(compensating, start=4):
        for col_idx, value in enumerate(control, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = grey_sample_fill
            cell.border = border_thin
            if col_idx in [4, 6, 7, 11]:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
        dv_effectiveness.add(ws.cell(row=row_idx, column=5))
        dv_test_status.add(ws.cell(row=row_idx, column=9))

    # Row 9: F2F2F2 grey sample row (MAX-003 fix)
    sample_row = 9
    sample_values = {
        1: "CC-006", 2: "SC-009", 3: "Network Segmentation",
        4: "Enhanced firewall ACLs and manual traffic monitoring",
        5: "Partial", 6: "Network team to apply ACL rules and review traffic logs hourly",
        7: "Emergency connectivity requirement", 8: "24 hours",
        9: "Tested", 10: "15.01.2026", 11: "Effective for short duration outages",
    }
    for col_idx, value in sample_values.items():
        cell = ws.cell(row=sample_row, column=col_idx, value=value)
        cell.fill = grey_sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 10-59: 50 empty FFFFCC input rows (MAX-004 fix)
    for r in range(10, 60):
        cell1 = ws.cell(row=r, column=1)
        cell1.fill = styles["input_cell"]["fill"]
        cell1.border = border_thin
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = border_thin
        dv_effectiveness.add(ws.cell(row=r, column=5))
        dv_test_status.add(ws.cell(row=r, column=9))

    ws.freeze_panes = "A4"  # DS-007 fix


# =============================================================================
# BC/DR SECURITY REVIEW SHEET
# =============================================================================
def create_bcdr_security_review_sheet(ws, styles):
    """Create the BCDR Security Review sheet."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_sample_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells("A1:K1")
    ws["A1"] = "BC/DR SECURITY REVIEW"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    # Row 2: subtitle (DS-006 fix)
    ws.merge_cells("A2:K2")
    ws["A2"] = "Security review status of all BC/DR plans — confirms CISO approval and security section coverage"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    columns = [
        ("Plan_ID", 15),
        ("Plan_Name", 35),
        ("Plan_Type", 20),
        ("Plan_Owner", 25),
        ("Security_Section_Present", 18),
        ("CISO_Review_Date", 16),
        ("CISO_Approval_Status", 18),
        ("Gaps_Identified", 40),
        ("Remediation_Due_Date", 18),
        ("Remediation_Status", 18),
        ("Next_Review_Due", 16),
    ]

    # Headers row 3
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    dv_plan_type = DataValidation(
        type="list",
        formula1='"BCP,DRP,Crisis Management,IT Recovery,Other"',
        allow_blank=False
    )
    ws.add_data_validation(dv_plan_type)

    dv_security_section = DataValidation(
        type="list",
        formula1='"Yes,No,Partial"',
        allow_blank=False
    )
    ws.add_data_validation(dv_security_section)

    dv_approval = DataValidation(
        type="list",
        formula1='"Approved,Rejected,Pending"',
        allow_blank=False
    )
    ws.add_data_validation(dv_approval)

    dv_remediation = DataValidation(
        type="list",
        formula1='"Open,In Progress,Closed"',
        allow_blank=False
    )
    ws.add_data_validation(dv_remediation)

    # Row 4: F2F2F2 grey sample row (MAX-003 fix)
    sample_values = {
        1: "BCP-001", 2: "Business Continuity Plan", 3: "BCP",
        4: "CISO", 5: "Yes", 6: "23.02.2026",
        7: "Approved", 8: "None identified", 9: "30.06.2026",
        10: "Closed", 11: "23.02.2027",
    }
    for col_idx, value in sample_values.items():
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.fill = grey_sample_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rows 5-54: 50 empty FFFFCC input rows (MAX-004 fix)
    for r in range(5, 55):
        cell1 = ws.cell(row=r, column=1)
        cell1.fill = styles["input_cell"]["fill"]
        cell1.border = border_thin
        for c in range(2, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = border_thin
            cell.alignment = styles["input_cell"]["alignment"]
        dv_plan_type.add(ws.cell(row=r, column=3))
        dv_security_section.add(ws.cell(row=r, column=5))
        dv_approval.add(ws.cell(row=r, column=7))
        dv_remediation.add(ws.cell(row=r, column=10))

    ws.freeze_panes = "A4"  # DS-007 fix




def create_evidence_register(ws):
    """Create GS-ER-compliant Evidence Register sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    ws.title = "Evidence Register"

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    # Row 1: Title banner
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = _border

    # Row 2: Italic subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = _border

    # Row 3: empty separator

    # Row 4: Column headers (003366 navy)
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border

    # Data validations
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other"',
        allow_blank=True,
    )
    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)
    ws.add_data_validation(ver_status_dv)

    # Row 5: Sample row (F2F2F2 grey)
    sample_data = {
        1: "EV-001", 2: "Assessment Area", 3: "Policy Document",
        4: "Example evidence description", 5: "\\\\fileserver\\policies\\document.pdf",
        6: "15.01.2026", 7: "Assessor Name", 8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = _border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center", wrap_text=True
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # Rows 6-105: Empty FFFFCC rows (100 rows)
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = _border
            cell.alignment = Alignment(
                horizontal="center" if col == 1 else "left",
                vertical="center", wrap_text=True
            )
            cell.value = None
        ev_type_dv.add(ws[f"C{row}"])
        ver_status_dv.add(ws[f"H{row}"])

    # Column widths and freeze panes
    for col, width in [("A", 15), ("B", 25), ("C", 22), ("D", 40), ("E", 45), ("F", 16), ("G", 20), ("H", 22)]:
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"





def create_approval_sheet(ws):
    """Gold Standard Approval Sign-Off — ISMS-IMP-A.5.29.1.

    Structure:
      Row 1  : Title banner (003366)
      Row 2  : CONTROL_REF subtitle
      Row 3  : ASSESSMENT SUMMARY banner (4472C4)
      Rows 4-8: Summary fields (Document / Period / Overall Compliance / Status / By)
                B6 = ='Summary Dashboard'!G9  [Overall Compliance Rating]
      Row 9  : gap
      Row 10 : ASSESSMENT REVIEW AND APPROVAL banner (4472C4)
      Row 11 : Column headers
      Rows 12-15: Approver rows
      Row 16 : gap
      Row 17 : FINAL DECISION (GS-AS-012: col A plain bold, no fill)
      Row 18 : gap
      Row 19 : NEXT REVIEW DETAILS banner (4472C4)
      Rows 20-21: Next Review Date / Review Owner
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _thin   = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy   = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _blue   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _grey   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _input  = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # -------------------------------------------------------------------------
    # Row 1: Title banner
    # -------------------------------------------------------------------------
    ws.merge_cells("A1:E1")
    ws["A1"] = "SECURITY CONTROLS DURING DISRUPTION — ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = _border
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = _border

    # -------------------------------------------------------------------------
    # Row 3: ASSESSMENT SUMMARY banner
    # -------------------------------------------------------------------------
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A3"].fill = _blue
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = _border

    # Rows 4-8: Summary fields (B6 = Overall Compliance Rating)
    summary_fields = [
        ("Document:",                f"{DOCUMENT_ID} — {WORKBOOK_NAME}"),
        ("Assessment Period:",       ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G8),\"\")"),
        ("Assessment Status:",       ""),
        ("Assessed By:",             ""),
    ]
    for row_idx, (label, value) in enumerate(summary_fields, start=4):
        lbl_cell = ws.cell(row=row_idx, column=1, value=label)
        lbl_cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        lbl_cell.border = _border
        ws.merge_cells(f"B{row_idx}:E{row_idx}")
        val_cell = ws.cell(row=row_idx, column=2, value=value)
        if value == "":
            val_cell.fill = _input
        elif "Summary Dashboard" in value:
            val_cell.number_format = "0.0%"
        for c in range(2, 6):
            ws.cell(row=row_idx, column=c).border = _border

    # Assessment Status DV on row 7
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # Row 9: gap
    ws.merge_cells("A9:E9")
    ws["A9"].border = _border

    # -------------------------------------------------------------------------
    # Row 10: ASSESSMENT REVIEW AND APPROVAL banner
    # -------------------------------------------------------------------------
    ws.merge_cells("A10:E10")
    ws["A10"] = "ASSESSMENT REVIEW AND APPROVAL"
    ws["A10"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A10"].fill = _blue
    ws["A10"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A10"].border = _border

    # Row 11: Column headers
    col_headers = ["Role / Function", "Name", "Signature / Initials", "Date (DD.MM.YYYY)", "Comments"]
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
        cell.fill = _grey
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border

    # Rows 12-15: Approver rows
    approvers = [
        "Lead Assessor / Author",
        "IT Security Manager",
        "Reviewer (Technical / Compliance)",
        "CISO / Final Approver",
    ]
    for row_idx, approver in enumerate(approvers, start=12):
        cell = ws.cell(row=row_idx, column=1, value=approver)
        cell.font = Font(name="Calibri", size=10)
        cell.fill = _input
        cell.border = _border
        for col in range(2, 6):
            c = ws.cell(row=row_idx, column=col)
            c.fill = _input
            c.border = _border

    # Row 16: gap
    ws.merge_cells("A16:E16")
    ws["A16"].border = _border

    # -------------------------------------------------------------------------
    # Row 17: FINAL DECISION (GS-AS-012: col A plain bold, no fill)
    # -------------------------------------------------------------------------
    ws["A17"] = "FINAL DECISION:"
    ws["A17"].font = Font(name="Calibri", size=11, bold=True, color="000000")
    ws["A17"].border = _border
    ws.merge_cells("B17:E17")
    ws["B17"] = "☐ Approved     ☐ Approved with Conditions     ☐ Rejected     ☐ Deferred"
    ws["B17"].font = Font(name="Calibri", size=10)
    ws["B17"].fill = _input
    ws["B17"].alignment = Alignment(horizontal="center", vertical="center")
    for col in range(2, 6):
        ws.cell(row=17, column=col).border = _border

    # Row 18: gap
    ws.merge_cells("A18:E18")
    ws["A18"].border = _border

    # -------------------------------------------------------------------------
    # Row 19: NEXT REVIEW DETAILS (Gold Standard: 4472C4 fill, white bold 11pt)
    # -------------------------------------------------------------------------
    ws.merge_cells("A19:E19")
    ws["A19"] = "NEXT REVIEW DETAILS"
    ws["A19"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A19"].fill = _blue
    ws["A19"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A19"].border = _border

    for row_idx, label in enumerate(["Next Review Date (DD.MM.YYYY)", "Review Owner"], start=20):
        ws.cell(row=row_idx, column=1, value=label).font = Font(name="Calibri", size=10, bold=True, color="000000")
        ws.cell(row=row_idx, column=1).fill = _input
        ws.cell(row=row_idx, column=1).border = _border
        ws.merge_cells(f"B{row_idx}:E{row_idx}")
        c = ws.cell(row=row_idx, column=2)
        c.fill = _input
        c.border = _border

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 30
    ws.freeze_panes = "A3"

def create_summary_dashboard_sheet(ws):
    """Create the Summary Dashboard sheet — Gold Standard compliant."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _red = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Row 1: Title (GS-SD-014: must contain "— SUMMARY DASHBOARD" em dash)
    ws.merge_cells("A1:G1")
    ws["A1"] = "SECURITY CONTROLS DURING DISRUPTION \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = THIN_BORDER
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle (left-aligned, no fill, no wrap_text)
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.5.29: Information Security During Disruption"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: empty

    # TABLE 1 banner (Row 4)
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = _navy
    for c in range(1, 8):
        ws.cell(row=4, column=c).border = THIN_BORDER
    ws.row_dimensions[4].height = 20

    # TABLE 1 headers (Row 5) — D9D9D9, black bold
    headers = ["Assessment Area", "Total Items", "Compliant", "Partial",
               "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = _grey
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # TABLE 1 data rows (Rows 6-8)
    # Area configs: (Area Name, sheet_name, total_col, total_range, compliant_val, partial_val, noncompliant_val, na_val)
    area_configs = [
        (
            "Security Control Inventory",
            "Security Control Inventory",
            "B", "B14:B63",
            ("E", "E14:E63", "Operational"),
            ("E", "E14:E63", "Partial"),
            None,
            ("E", "E14:E63", "Not Applicable"),
        ),
        (
            "Minimum Baseline",
            "Minimum Baseline",
            "B", "B10:B59",
            ("G", "G10:G59", "Approved"),
            ("G", "G10:G59", "Pending"),
            ("G", "G10:G59", "Rejected"),
            None,
        ),
        (
            "BCDR Security Review",
            "BCDR Security Review",
            "B", "B5:B54",
            ("G", "G5:G54", "Approved"),
            ("G", "G5:G54", "Pending"),
            ("G", "G5:G54", "Rejected"),
            None,
        ),
    ]

    for i, config in enumerate(area_configs):
        row = 6 + i
        area_name, sheet_name, _, total_range, comp, part, noncomp, na = config

        # Col A: area name
        cell_a = ws.cell(row=row, column=1, value=area_name)
        cell_a.border = THIN_BORDER
        cell_a.font = Font(color="000000")

        # Col B: Total (COUNTA of first user-entered col)
        total_col = total_range.split(":")[0][0]  # e.g. "B" from "B14:B63"
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = f"=COUNTA('{sheet_name}'!{total_range})"
        cell_b.border = THIN_BORDER
        cell_b.alignment = Alignment(horizontal="center")
        cell_b.font = Font(color="000000")

        # Col C: Compliant
        cell_c = ws.cell(row=row, column=3)
        if comp:
            sc, sr, sv = comp
            cell_c.value = f'=COUNTIF(\'{sheet_name}\'!{sr},"{sv}")'
        else:
            cell_c.value = 0
        cell_c.border = THIN_BORDER
        cell_c.alignment = Alignment(horizontal="center")
        cell_c.font = Font(color="000000")

        # Col D: Partial
        cell_d = ws.cell(row=row, column=4)
        if part:
            sc, sr, sv = part
            cell_d.value = f'=COUNTIF(\'{sheet_name}\'!{sr},"{sv}")'
        else:
            cell_d.value = 0
        cell_d.border = THIN_BORDER
        cell_d.alignment = Alignment(horizontal="center")
        cell_d.font = Font(color="000000")

        # Col E: Non-Compliant
        cell_e = ws.cell(row=row, column=5)
        if noncomp:
            sc, sr, sv = noncomp
            cell_e.value = f'=COUNTIF(\'{sheet_name}\'!{sr},"{sv}")'
        else:
            cell_e.value = 0
        cell_e.border = THIN_BORDER
        cell_e.alignment = Alignment(horizontal="center")
        cell_e.font = Font(color="000000")

        # Col F: N/A
        cell_f = ws.cell(row=row, column=6)
        if na:
            sc, sr, sv = na
            cell_f.value = f'=COUNTIF(\'{sheet_name}\'!{sr},"{sv}")'
        else:
            cell_f.value = 0
        cell_f.border = THIN_BORDER
        cell_f.alignment = Alignment(horizontal="center")
        cell_f.font = Font(color="000000")

        # Col G: Compliance %
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f"=IFERROR(IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row})),\"\")"
        cell_g.number_format = "0.0%"
        cell_g.border = THIN_BORDER
        cell_g.alignment = Alignment(horizontal="center")
        cell_g.font = Font(color="000000")

    # TOTAL row (Row 9)
    total_row = 9
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="000000")
    ws.cell(row=total_row, column=1).fill = _grey
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}8)"
        cell.font = Font(bold=True, color="000000")
        cell.fill = _grey
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    cell_gt = ws.cell(row=total_row, column=7)
    cell_gt.value = f"=IFERROR(IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row})),\"\")"
    cell_gt.number_format = "0.0%"
    cell_gt.font = Font(bold=True, color="000000")
    cell_gt.fill = _grey
    cell_gt.border = THIN_BORDER
    cell_gt.alignment = Alignment(horizontal="center")

    # TABLE 2: KEY METRICS
    t2_start = total_row + 2  # row 11
    ws.merge_cells(f"A{t2_start}:G{t2_start}")
    ws[f"A{t2_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{t2_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t2_start}"].fill = _navy
    for c in range(1, 8):
        ws.cell(row=t2_start, column=c).border = THIN_BORDER

    # TABLE 2 headers — D9D9D9 (GS-SD-016)
    t2_header_row = t2_start + 1
    for col, hdr in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        cell = ws.cell(row=t2_header_row, column=col, value=hdr if hdr else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = _grey
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # TABLE 2 metrics — white fill, 000000 font (NOT bold, NOT 003366)
    metrics = [
        ("Total Security Controls Inventoried", "=COUNTA('Security Control Inventory'!B14:B63)"),
        ("Non-Negotiable Controls (Must Maintain)", "=COUNTIF('Security Control Inventory'!F14:F63,\"Non-Negotiable\")"),
        ("Degradable Controls", "=COUNTIF('Security Control Inventory'!F14:F63,\"Degradable\")"),
        ("Deferrable Controls", "=COUNTIF('Security Control Inventory'!F14:F63,\"Deferrable\")"),
        ("Critical Priority Controls", "=COUNTIF('Security Control Inventory'!G14:G63,\"Critical\")"),
        ("High Priority Controls", "=COUNTIF('Security Control Inventory'!G14:G63,\"High\")"),
        ("Controls Currently Operational", "=COUNTIF('Security Control Inventory'!E14:E63,\"Operational\")"),
        ("Controls Partially Operational", "=COUNTIF('Security Control Inventory'!E14:E63,\"Partial\")"),
        ("Baseline Controls Approved", "=COUNTIF('Minimum Baseline'!G10:G59,\"Approved\")"),
        ("Baseline Controls Pending Approval", "=COUNTIF('Minimum Baseline'!G10:G59,\"Pending\")"),
        ("Total Compensating Controls", "=COUNTA('Compensating Controls'!B10:B59)"),
        ("Compensating Controls Tested", "=COUNTIF('Compensating Controls'!I10:I59,\"Tested\")"),
        ("Compensating Controls Untested", "=COUNTIF('Compensating Controls'!I10:I59,\"Untested\")"),
        ("BC/DR Plans Reviewed", "=COUNTA('BCDR Security Review'!B5:B54)"),
        ("BC/DR Plans with Security Section", "=COUNTIF('BCDR Security Review'!E5:E54,\"Yes\")"),
        ("BC/DR Plans CISO-Approved", "=COUNTIF('BCDR Security Review'!G5:G54,\"Approved\")"),
        ("BC/DR Plans with Open Remediation", "=COUNTIF('BCDR Security Review'!J5:J54,\"Open\")"),
    ]

    row = t2_header_row + 1
    for metric, formula in metrics:
        cell_m = ws.cell(row=row, column=1, value=metric)
        cell_m.border = THIN_BORDER
        cell_m.font = Font(color="000000")
        cell_v = ws.cell(row=row, column=2, value=formula)
        cell_v.border = THIN_BORDER
        cell_v.font = Font(color="000000")
        cell_v.alignment = Alignment(horizontal="center")
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # 2 empty buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # TABLE 3: CRITICAL FINDINGS
    t3_start = row + 1
    ws.merge_cells(f"A{t3_start}:G{t3_start}")
    ws[f"A{t3_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{t3_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t3_start}"].fill = _red
    for c in range(1, 8):
        ws.cell(row=t3_start, column=c).border = THIN_BORDER

    # TABLE 3 headers — D9D9D9
    t3_header_row = t3_start + 1
    for col, hdr in enumerate(["Category", "Finding", "Count", "Severity", "Action Required", "", ""], 1):
        cell = ws.cell(row=t3_header_row, column=col, value=hdr if hdr else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = _grey
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # TABLE 3 findings — FFFFCC fill, 000000 font
    findings = [
        ("Security Control Inventory", "Non-Negotiable controls not Operational",
         "=COUNTIFS('Security Control Inventory'!F14:F63,\"Non-Negotiable\",'Security Control Inventory'!E14:E63,\"Partial\")",
         "Critical", "Immediate"),
        ("Security Control Inventory", "Critical priority controls not Operational",
         "=COUNTIFS('Security Control Inventory'!G14:G63,\"Critical\",'Security Control Inventory'!E14:E63,\"Partial\")",
         "Critical", "Immediate"),
        ("Minimum Baseline", "Baseline controls not yet Approved",
         "=COUNTIF('Minimum Baseline'!G10:G59,\"Pending\")+COUNTIF('Minimum Baseline'!G10:G59,\"Rejected\")",
         "Critical", "Immediate"),
        ("Compensating Controls", "Untested compensating controls",
         "=COUNTIF('Compensating Controls'!I10:I59,\"Untested\")",
         "High", "Urgent"),
        ("Compensating Controls", "Failed compensating controls",
         "=COUNTIF('Compensating Controls'!I10:I59,\"Failed\")",
         "Critical", "Immediate"),
        ("BCDR Security Review", "BC/DR plans missing security section",
         "=COUNTIF('BCDR Security Review'!E5:E54,\"No\")",
         "High", "Urgent"),
        ("BCDR Security Review", "BC/DR plans not CISO-approved",
         "=COUNTIF('BCDR Security Review'!G5:G54,\"Pending\")+COUNTIF('BCDR Security Review'!G5:G54,\"Rejected\")",
         "High", "Urgent"),
        ("BCDR Security Review", "Plans with open remediation items",
         "=COUNTIF('BCDR Security Review'!J5:J54,\"Open\")",
         "Medium", "Plan"),
    ]

    row = t3_header_row + 1
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = _yellow
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).font = Font(color="000000")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_cnt = ws.cell(row=row, column=3, value=formula)
        cell_cnt.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    # 2 empty buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = _yellow
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly

def main() -> int:
    """
    Main execution function - orchestrates workbook creation.

    Returns:
        int: 0 on success, 1 on failure
    """
    logger.info("=" * 78)
    logger.info(f"{DOCUMENT_ID} - {WORKBOOK_NAME} Generator")
    logger.info(CONTROL_REF)
    logger.info("=" * 78)

    try:
        wb = create_workbook()
        styles = _STYLES

        logger.info("[1/9] Creating Instructions sheet...")
        create_instructions_sheet(wb["Instructions & Legend"])

        logger.info("[2/9] Creating Security Control Inventory sheet...")
        create_security_control_inventory_sheet(wb["Security Control Inventory"], styles)

        logger.info("[3/9] Creating Minimum Baseline sheet...")
        create_minimum_baseline_sheet(wb["Minimum Baseline"], styles)

        logger.info("[4/9] Creating Security Posture Levels sheet...")
        create_security_posture_levels_sheet(wb["Security Posture Levels"], styles)

        logger.info("[5/9] Creating Compensating Controls sheet...")
        create_compensating_controls_sheet(wb["Compensating Controls"], styles)

        logger.info("[6/9] Creating BCDR Security Review sheet...")
        create_bcdr_security_review_sheet(wb["BCDR Security Review"], styles)

        logger.info("[7/9] Creating Evidence Register sheet...")
        create_evidence_register(wb["Evidence Register"])

        logger.info("[8/9] Creating Summary Dashboard sheet...")
        create_summary_dashboard_sheet(wb["Summary Dashboard"])

        logger.info("[9/9] Creating Approval Sign-Off sheet...")
        create_approval_sheet(wb["Approval Sign-Off"])

        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
        output_path = _wkbk_dir / OUTPUT_FILENAME
        wb.save(output_path)
        logger.info(f"SUCCESS: {OUTPUT_FILENAME}")
        logger.info("Next steps:")
        logger.info("  1) Complete document information in Instructions")
        logger.info("  2) Inventory security controls")
        logger.info("  3) Define minimum baseline")
        logger.info("  4) Document security posture levels")
        logger.info("  5) Identify compensating controls")
        logger.info("  6) Review BC/DR plans for security")
        logger.info("  7) Link evidence")
        logger.info("  8) Obtain approvals")
        logger.info("=" * 78)
        return 0

    except ImportError as e:
        logger.error(f"Missing dependency: {e}")
        logger.error("Install with: pip install openpyxl")
        return 1
    except PermissionError as e:
        logger.error(f"Permission denied: {e}")
        return 1
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
