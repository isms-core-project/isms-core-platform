#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.1-2-6.1-2.S1 - Policy Framework Assessment Workbook Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.1-2-6.1-2: Information Security Policies and Organisation
Assessment Domain 1 of 4: Policy Framework Assessment Workbook

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific information security policies and organisation infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Policy inventory scope and mandatory policy categories (match your organisation)
2. Role definition framework and responsibility assignment methodology
3. Screening procedure requirements and verification standards (adapt to jurisdiction)
4. Employment contract security clause requirements (legal review required)
5. Policy governance cycle including approval authorities and review owners

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.1-2-6.1-2 Information Security Policies and Organisation Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
information security policies and organisation controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Policy Framework Assessment Workbook under ISO 27001:2022 Controls A.5.1, A.5.2, A.6.1, and A.6.2. Supports evidence-based evaluation of policy governance, role accountability, personnel screening, and employment framework compliance.

**Assessment Scope:**
- Policy inventory completeness and lifecycle compliance
- Role and responsibility assignment accuracy and coverage
- Personnel screening and vetting procedure adherence
- Employment contract security clause documentation
- Policy governance and approval workflow effectiveness
- Communication and acknowledgment tracking completeness
- Evidence collection for HR, governance, and compliance audits

**Generated Workbook Structure:**
1. Policy Inventory
2. Lifecycle Compliance
3. Governance Assessment
4. Classification Review
5. Communication Tracking
6. Repository Assessment
7. Gap Analysis
8. Evidence Register
9. Action Items
10. Approval Sign-Off

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 4 domains covering Information Security Policies and Organisation controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a5_1_2_6_1_2_s1_policy_framework.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a5_1_2_6_1_2_s1_policy_framework.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a5_1_2_6_1_2_s1_policy_framework.py --date 20250115

Output:
    File: ISMS-IMP-A.5.1-2-6.1-2.S1_Policy_Framework_Assessment_Workbook_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.1-2-6.1-2
Assessment Domain:    1 of 4 (Policy Framework Assessment Workbook)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.1-2-6.1-2: Information Security Policies and Organisation Policy (Governance)
    - ISMS-IMP-A.5.1-2-6.1-2.S1: Policy Framework Assessment Workbook (Domain 1)
    - ISMS-IMP-A.5.1-2-6.1-2.S2: Roles Responsibilities Assessment Workbook (Domain 2)
    - ISMS-IMP-A.5.1-2-6.1-2.S3: Screening Vetting Assessment Workbook (Domain 3)
    - ISMS-IMP-A.5.1-2-6.1-2.S4: Employment Contract Assessment (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.1-2-6.1-2.S1 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive information security policies and organisation details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review policy inventory, role definitions, and screening procedures annually or when organisational restructuring occurs, regulatory requirements change, or HR processes are updated.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.1-2-6.1-2.S1"
WORKBOOK_NAME = "Policy Framework Assessment Workbook"
CONTROL_ID = "A.5.1-2-6.1-2"
CONTROL_NAME = "Information Security Policies and Organisation"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# WKBK output directory
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ==============================================================================
# STYLE DEFINITIONS
# ==============================================================================
def setup_styles():
    """
    Define consistent cell styles used throughout the workbook.
    
    Returns:
        dict: Dictionary of style objects for different cell types
    """
    styles = {}
    
    # Header styles
    styles['header_main'] = {
        'font': Font(name='Calibri', size=18, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }
    
    styles['header_sub'] = {
        'font': Font(name='Calibri', size=12, bold=False, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }
    
    styles['section_header'] = {
        'font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }
    
    styles['column_header'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }
    
    # Data cell styles
    styles['input_cell'] = {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),  # Yellow
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }
    
    styles['calculated_cell'] = {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),  # Grey (standard palette)
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }
    
    styles['label_cell'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),  # Gray
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }
    
    # Status fills (for conditional formatting reference)
    styles['status_compliant'] = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
    styles['status_partial'] = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')  # Yellow
    styles['status_non_compliant'] = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red
    
    return styles


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================


_STYLES = setup_styles()
def apply_cell_style(cell, style_dict):
    """
    Apply a style dictionary to a cell.
    
    Args:
        cell: openpyxl cell object
        style_dict: Dictionary containing style attributes
    """
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']


def set_column_widths(ws, widths_dict):
    """
    Set column widths for a worksheet.
    
    Args:
        ws: Worksheet object
        widths_dict: Dictionary mapping column letters to widths
    """
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width


def create_data_validation(formula_or_list, allow_blank=True, error_title="Invalid Entry", 
                           error_message="Please select from dropdown"):
    """
    Create a data validation object for dropdown lists.
    
    Args:
        formula_or_list: Either a string formula or list of values
        allow_blank: Whether to allow blank cells
        error_title: Title for error message
        error_message: Error message text
    
    Returns:
        DataValidation object
    """
    if isinstance(formula_or_list, list):
        dv = DataValidation(type="list", formula1=f'"{",".join(formula_or_list)}"', 
                           allow_blank=allow_blank)
    else:
        dv = DataValidation(type="list", formula1=formula_or_list, allow_blank=allow_blank)
    
    dv.error = error_message
    dv.errorTitle = error_title
    dv.showErrorMessage = True
    return dv


def merge_and_style(ws, range_string, value, style_dict):
    """
    Merge cells and apply style.
    
    Args:
        ws: Worksheet object
        range_string: Range to merge (e.g., 'A1:K1')
        value: Value to write
        style_dict: Style dictionary to apply
    """
    ws.merge_cells(range_string)
    cell = ws[range_string.split(':')[0]]
    cell.value = value
    apply_cell_style(cell, style_dict)


# ==============================================================================
# SHEET CREATION FUNCTIONS
# ==============================================================================


def create_policy_inventory_sheet(wb, styles):
    """
    Sheet 2: Policy Inventory
    
    Master policy inventory with full metadata.
    150 rows for policies.
    """
    ws = wb.create_sheet("Policy Inventory")
    ws.sheet_view.showGridLines = False
    
    # --- Header ---
    merge_and_style(ws, 'A1:T1', 'POLICY INVENTORY', styles['header_main'])
    merge_and_style(ws, 'A2:T2',
                   'Master list of all information security policies with metadata',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35
    
    # --- Column Headers (Row 3) ---
    headers = [
        ('A', 'Policy ID', 20),
        ('B', 'Policy Title', 35),
        ('C', 'Policy Hierarchy Tier', 15),
        ('D', 'ISO Control Mapping', 20),
        ('E', 'Policy Owner', 25),
        ('F', 'Policy Approver', 25),
        ('G', 'Current Version', 12),
        ('H', 'Version Date', 12),
        ('I', 'Approval Date', 12),
        ('J', 'Last Review Date', 12),
        ('K', 'Next Review Date', 12),
        ('L', 'Review Frequency', 15),
        ('M', 'Policy Status', 15),
        ('N', 'Policy Classification', 15),
        ('O', 'Acknowledgment Required', 15),
        ('P', 'Acknowledgment Rate', 12),
        ('Q', 'Repository Location', 40),
        ('R', 'Related Documents', 30),
        ('S', 'Gap Identified', 15),
        ('T', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}3'] = header_text
        apply_cell_style(ws[f'{col_letter}3'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Sample Row (Row 4) — Step 2.4 ---
    _pi_smp_fill = PatternFill('solid', fgColor='F2F2F2')
    _pi_smp_font = Font(name='Calibri', size=10, italic=True, color='808080')
    _pi_smp_bdr = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    _pi_sample = {
        'A': 'POL-EXAMPLE', 'B': 'Information Security Policy (Example)',
        'C': 'Tier-1-Master', 'D': 'A.5.1, A.5.2',
        'E': 'Chief Information Security Officer', 'F': 'Chief Executive Officer',
        'G': 'v2.0', 'H': '01.01.2025', 'I': '15.01.2025',
        'J': '15.01.2026', 'K': '15.01.2027', 'L': 'Annual',
        'M': 'Active', 'N': 'Internal', 'O': 'Yes', 'P': 0.92,
        'Q': 'SharePoint/ISMS Repository', 'R': 'ISMS-IMP-A.5.1 User Guide',
        'S': 'No-Gap', 'T': 'Example entry \u2014 replace with your policy details',
    }
    for col_letter in [c[0] for c in headers]:
        ws[f'{col_letter}4'].value = _pi_sample.get(col_letter, '')
        ws[f'{col_letter}4'].fill = _pi_smp_fill
        ws[f'{col_letter}4'].font = _pi_smp_font
        ws[f'{col_letter}4'].border = _pi_smp_bdr
    ws['P4'].number_format = '0%'

    # --- Data Rows (5-54) — 50 empty FFFFCC rows (gold standard) ---
    for row in range(5, 55):
        for col_letter in [c[0] for c in headers]:
            cell = ws[f'{col_letter}{row}']
            apply_cell_style(cell, styles['input_cell'])

    # --- Data Validation ---
    # Tier dropdown
    tier_dv = create_data_validation(['Tier-1-Master', 'Tier-2-Domain', 'Tier-3-Topic'])
    ws.add_data_validation(tier_dv)
    tier_dv.add(f'C5:C54')

    # Review Frequency dropdown
    freq_dv = create_data_validation(['Annual', 'Biennial', 'Quarterly', 'Triggered-Only'])
    ws.add_data_validation(freq_dv)
    freq_dv.add(f'L5:L54')

    # Status dropdown
    status_dv = create_data_validation(['Active', 'Draft', 'Under-Review', 'Retired', 'Superseded'])
    ws.add_data_validation(status_dv)
    status_dv.add(f'M5:M54')

    # Classification dropdown
    class_dv = create_data_validation(['Internal', 'Confidential', 'Public'])
    ws.add_data_validation(class_dv)
    class_dv.add(f'N5:N54')

    # Acknowledgment Required dropdown
    ack_dv = create_data_validation(['Yes', 'No', 'Role-Specific'])
    ws.add_data_validation(ack_dv)
    ack_dv.add(f'O5:O54')

    # Gap Identified dropdown
    gap_dv = create_data_validation(['No-Gap', 'Minor-Gap', 'Significant-Gap', 'Critical-Gap'])
    ws.add_data_validation(gap_dv)
    gap_dv.add(f'S5:S54')

    # Date formatting
    for col in ['H', 'I', 'J', 'K']:
        for row in range(5, 55):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # Percentage formatting for Acknowledgment_Rate
    for row in range(5, 55):
        ws[f'P{row}'].number_format = '0%'

    # Freeze panes at A5 (rows 1-4 frozen: title, subtitle, headers, sample)
    ws.freeze_panes = 'A4'

    # Define named range for Policy_ID_List
    policy_id_range = DefinedName(name='Policy_ID_List', attr_text="\'Policy Inventory\'!$A$5:$A$54")
    wb.defined_names.add(policy_id_range)
    
    # Sheet protection removed (SRC-018)


def create_lifecycle_compliance_sheet(wb, styles):
    """
    Sheet 3: Lifecycle Compliance
    
    Verify policy lifecycle stages (creation, approval, review, update).
    """
    ws = wb.create_sheet("Lifecycle Compliance")
    ws.sheet_view.showGridLines = False
    
    # --- Header ---
    merge_and_style(ws, 'A1:N1', 'LIFECYCLE COMPLIANCE ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:N2',
                   'Verification of policy creation, approval, publication, review, and update processes',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35
    
    # --- Column Headers ---
    headers = [
        ('A', 'Policy ID', 20),
        ('B', 'Policy Title', 35),
        ('C', 'Creation Process Followed', 20),
        ('D', 'Approval Valid', 15),
        ('E', 'Approval Documentation', 20),
        ('F', 'Publication Status', 20),
        ('G', 'Review Schedule Defined', 20),
        ('H', 'Review Status', 20),
        ('I', 'Last Review Evidence', 20),
        ('J', 'Version Control Practice', 15),
        ('K', 'Sunset Process', 15),
        ('L', 'Lifecycle Compliance Rating', 20),
        ('M', 'Gap Description', 40),
        ('N', 'Evidence Reference', 30)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)

    # --- Sample Row (Row 5, F2F2F2 grey) ---
    _smp_fill = PatternFill('solid', fgColor='F2F2F2')
    _smp_font = Font(name='Calibri', size=10, italic=True, color='808080')
    _smp_bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
    _lc_sample = {
        'A': 'POL-EXAMPLE', 'B': 'Information Security Policy (Example)',
        'C': 'Yes', 'D': 'Yes', 'E': 'Complete', 'F': 'Published',
        'G': 'Yes', 'H': 'Current', 'I': 'Yes', 'J': 'Good',
        'K': 'N/A', 'L': 'Compliant',
        'M': '', 'N': 'Example — replace with actual lifecycle data',
    }
    for col_letter in [c[0] for c in headers]:
        ws[f'{col_letter}5'].value = _lc_sample.get(col_letter, '')
        ws[f'{col_letter}5'].fill = _smp_fill
        ws[f'{col_letter}5'].font = _smp_font
        ws[f'{col_letter}5'].border = _smp_bdr

    # --- Data Rows (6-155) with Auto-Population Formulas ---
    for row in range(6, 156):
        # Auto-populate Policy_ID from Policy Inventory
        ws[f'A{row}'] = f'=IF(ROW()-5<=COUNTA(\'Policy Inventory\'!$A:$A)-4,INDEX(\'Policy Inventory\'!$A:$A,ROW()-1),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # Auto-populate Policy_Title
        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},\'Policy Inventory\'!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])

        # Manual entry cells (C-G, I-K, M-N)
        for col in ['C', 'D', 'E', 'F', 'G', 'I', 'J', 'K', 'M', 'N']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # Review_Status (H) - auto-calculated
        ws[f'H{row}'] = (f'=IF(B{row}="","",IF(VLOOKUP(A{row},\'Policy Inventory\'!$A:$K,11,FALSE)>=TODAY(),"Current",'
                        f'IF(VLOOKUP(A{row},\'Policy Inventory\'!$A:$K,11,FALSE)>=TODAY()-30,"Overdue-<30-Days",'
                        f'IF(VLOOKUP(A{row},\'Policy Inventory\'!$A:$K,11,FALSE)>=TODAY()-90,"Overdue-30-90-Days","Overdue->90-Days"))))')
        apply_cell_style(ws[f'H{row}'], styles['calculated_cell'])

        # Lifecycle Compliance Rating (L) - auto-calculated
        ws[f'L{row}'] = (f'=IF(B{row}="","",IF(AND(D{row}="Yes",E{row}="Complete",F{row}="Published",'
                        f'OR(H{row}="Current",H{row}="Overdue-<30-Days"),I{row}="Yes",J{row}<>"Poor"),"Compliant",'
                        f'IF(OR(D{row}="No",D{row}="Expired",E{row}="Missing",F{row}="Not-Published",H{row}="Overdue->90-Days"),"Non-Compliant","Partial")))')
        apply_cell_style(ws[f'L{row}'], styles['calculated_cell'])

    # --- Data Validation ---
    creation_dv = create_data_validation(['Yes', 'Partial', 'No', 'Unknown'])
    ws.add_data_validation(creation_dv)
    creation_dv.add('C6:C155')

    approval_valid_dv = create_data_validation(['Yes', 'Partial', 'No', 'Expired'])
    ws.add_data_validation(approval_valid_dv)
    approval_valid_dv.add('D6:D155')

    approval_doc_dv = create_data_validation(['Complete', 'Incomplete', 'Missing'])
    ws.add_data_validation(approval_doc_dv)
    approval_doc_dv.add('E6:E155')

    pub_status_dv = create_data_validation(['Published', 'Not-Published', 'Partially-Accessible'])
    ws.add_data_validation(pub_status_dv)
    pub_status_dv.add('F6:F155')

    review_schedule_dv = create_data_validation(['Yes', 'No'])
    ws.add_data_validation(review_schedule_dv)
    review_schedule_dv.add('G6:G155')

    review_evidence_dv = create_data_validation(['Yes', 'Partial', 'No'])
    ws.add_data_validation(review_evidence_dv)
    review_evidence_dv.add('I6:I155')

    version_control_dv = create_data_validation(['Excellent', 'Good', 'Adequate', 'Poor'])
    ws.add_data_validation(version_control_dv)
    version_control_dv.add('J6:J155')

    sunset_dv = create_data_validation(['N/A', 'Yes', 'No'])
    ws.add_data_validation(sunset_dv)
    sunset_dv.add('K6:K155')

    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Sheet protection removed (SRC-018)


def create_governance_assessment_sheet(wb, styles):
    """
    Sheet 4: Governance Assessment
    
    Ownership, accountability, approval authority verification.
    """
    ws = wb.create_sheet("Governance Assessment")
    ws.sheet_view.showGridLines = False
    
    # --- Header ---
    merge_and_style(ws, 'A1:O1', 'GOVERNANCE ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:O2',
                   'Ownership, accountability, approval authority, and RACI verification',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35
    
    # --- Column Headers ---
    headers = [
        ('A', 'Policy ID', 20),
        ('B', 'Policy Title', 35),
        ('C', 'Policy Hierarchy Tier', 15),
        ('D', 'Owner Assigned', 15),
        ('E', 'Owner Name Role', 25),
        ('F', 'Owner Accountability Clear', 15),
        ('G', 'Approver Assigned', 15),
        ('H', 'Approver Name Role', 25),
        ('I', 'Approval Authority Appropriate', 20),
        ('J', 'RACI Defined', 15),
        ('K', 'Governance Documentation', 20),
        ('L', 'Escalation Path Clear', 15),
        ('M', 'Governance Compliance Rating', 20),
        ('N', 'Gap Description', 40),
        ('O', 'Evidence Reference', 30)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)

    # --- Sample Row (Row 5, F2F2F2 grey) ---
    _smp_fill = PatternFill('solid', fgColor='F2F2F2')
    _smp_font = Font(name='Calibri', size=10, italic=True, color='808080')
    _smp_bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
    _ga_sample = {
        'A': 'POL-EXAMPLE', 'B': 'Information Security Policy (Example)',
        'C': 'Tier-1-Master', 'D': 'Yes', 'E': 'Chief Information Security Officer',
        'F': 'Yes', 'G': 'Yes', 'H': 'Chief Executive Officer', 'I': 'Yes',
        'J': 'Yes', 'K': 'Complete', 'L': 'Yes', 'M': 'Compliant',
        'N': '', 'O': 'Example — replace with actual governance data',
    }
    for col_letter in [c[0] for c in headers]:
        ws[f'{col_letter}5'].value = _ga_sample.get(col_letter, '')
        ws[f'{col_letter}5'].fill = _smp_fill
        ws[f'{col_letter}5'].font = _smp_font
        ws[f'{col_letter}5'].border = _smp_bdr

    # --- Data Rows ---
    for row in range(6, 156):
        # Auto-populate columns
        ws[f'A{row}'] = f'=IF(ROW()-5<=COUNTA(\'Policy Inventory\'!$A:$A)-4,INDEX(\'Policy Inventory\'!$A:$A,ROW()-1),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},\'Policy Inventory\'!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])

        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},\'Policy Inventory\'!$A:$C,3,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])

        ws[f'D{row}'] = f'=IF(A{row}="","",IF(VLOOKUP(A{row},\'Policy Inventory\'!$A:$E,5,FALSE)<>"","Yes","No"))'
        apply_cell_style(ws[f'D{row}'], styles['calculated_cell'])

        ws[f'E{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},\'Policy Inventory\'!$A:$E,5,FALSE),"")'
        apply_cell_style(ws[f'E{row}'], styles['calculated_cell'])

        ws[f'G{row}'] = f'=IF(A{row}="","",IF(VLOOKUP(A{row},\'Policy Inventory\'!$A:$F,6,FALSE)<>"","Yes","No"))'
        apply_cell_style(ws[f'G{row}'], styles['calculated_cell'])

        ws[f'H{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},\'Policy Inventory\'!$A:$F,6,FALSE),"")'
        apply_cell_style(ws[f'H{row}'], styles['calculated_cell'])

        # Manual entry cells
        for col in ['F', 'I', 'J', 'K', 'L', 'N', 'O']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # Governance Compliance Rating (M) - auto-calculated
        ws[f'M{row}'] = (f'=IF(B{row}="","",IF(AND(D{row}="Yes",G{row}="Yes",I{row}="Yes",K{row}="Complete",L{row}="Yes"),"Compliant",'
                        f'IF(OR(D{row}="No",G{row}="No",I{row}="Insufficient",K{row}="Missing"),"Non-Compliant","Partial")))')
        apply_cell_style(ws[f'M{row}'], styles['calculated_cell'])

    # --- Data Validation ---
    owner_account_dv = create_data_validation(['Yes', 'Unclear', 'No'])
    ws.add_data_validation(owner_account_dv)
    owner_account_dv.add('F6:F155')

    approval_auth_dv = create_data_validation(['Yes', 'No', 'Elevated-Unnecessarily', 'Insufficient'])
    ws.add_data_validation(approval_auth_dv)
    approval_auth_dv.add('I6:I155')

    raci_dv = create_data_validation(['Yes', 'Partial', 'No', 'N/A'])
    ws.add_data_validation(raci_dv)
    raci_dv.add('J6:J155')

    gov_doc_dv = create_data_validation(['Complete', 'Partial', 'Missing'])
    ws.add_data_validation(gov_doc_dv)
    gov_doc_dv.add('K6:K155')

    escalation_dv = create_data_validation(['Yes', 'Unclear', 'No'])
    ws.add_data_validation(escalation_dv)
    escalation_dv.add('L6:L155')

    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Sheet protection removed (SRC-018)


def create_classification_review_sheet(wb, styles):
    """
    Sheet 5: Classification Review
    
    Policy classification appropriateness and access control verification.
    """
    ws = wb.create_sheet("Classification Review")
    ws.sheet_view.showGridLines = False
    
    # --- Header ---
    merge_and_style(ws, 'A1:L1', 'CLASSIFICATION REVIEW', styles['header_main'])
    merge_and_style(ws, 'A2:L2',
                   'Policy classification appropriateness and access control verification',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35
    
    # --- Column Headers ---
    headers = [
        ('A', 'Policy ID', 20),
        ('B', 'Policy Title', 35),
        ('C', 'Current Classification', 15),
        ('D', 'Classification Appropriate', 20),
        ('E', 'Content Sensitivity Assessment', 15),
        ('F', 'Access Controls Implemented', 20),
        ('G', 'Distribution Restrictions', 20),
        ('H', 'Classification Marking', 15),
        ('I', 'Classification Review Date', 15),
        ('J', 'Classification Compliance Rating', 20),
        ('K', 'Gap Description', 40),
        ('L', 'Evidence Reference', 30)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)

    # --- Sample Row (Row 5, F2F2F2 grey) ---
    _smp_fill = PatternFill('solid', fgColor='F2F2F2')
    _smp_font = Font(name='Calibri', size=10, italic=True, color='808080')
    _smp_bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
    _cr_sample = {
        'A': 'POL-EXAMPLE', 'B': 'Information Security Policy (Example)',
        'C': 'Internal', 'D': 'Yes', 'E': 'Medium',
        'F': 'Yes', 'G': 'Yes', 'H': 'Yes', 'I': '01.01.2026',
        'J': 'Compliant', 'K': '', 'L': 'Example — replace with actual classification data',
    }
    for col_letter in [c[0] for c in headers]:
        ws[f'{col_letter}5'].value = _cr_sample.get(col_letter, '')
        ws[f'{col_letter}5'].fill = _smp_fill
        ws[f'{col_letter}5'].font = _smp_font
        ws[f'{col_letter}5'].border = _smp_bdr

    # --- Data Rows (6-155) ---
    for row in range(6, 156):
        # Auto-populate columns
        ws[f'A{row}'] = f'=IF(ROW()-5<=COUNTA(\'Policy Inventory\'!$A:$A)-4,INDEX(\'Policy Inventory\'!$A:$A,ROW()-1),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},\'Policy Inventory\'!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])

        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},\'Policy Inventory\'!$A:$N,14,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])

        # Manual entry cells
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'K', 'L']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # Classification_Compliance_Rating (J) - auto-calculated
        ws[f'J{row}'] = (f'=IF(B{row}="","",IF(AND(D{row}="Yes",F{row}="Yes",H{row}="Yes"),"Compliant",'
                        f'IF(OR(D{row}="No",F{row}="No",H{row}="No"),"Non-Compliant","Partial")))')
        apply_cell_style(ws[f'J{row}'], styles['calculated_cell'])

    # Date formatting for Classification Review Date
    for row in range(6, 156):
        ws[f'I{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    class_approp_dv = create_data_validation(['Yes', 'No', 'Should-Be-Higher', 'Should-Be-Lower'])
    ws.add_data_validation(class_approp_dv)
    class_approp_dv.add('D6:D155')

    sensitivity_dv = create_data_validation(['None', 'Low', 'Medium', 'High'])
    ws.add_data_validation(sensitivity_dv)
    sensitivity_dv.add('E6:E155')

    access_controls_dv = create_data_validation(['Yes', 'Partial', 'No', 'Unknown'])
    ws.add_data_validation(access_controls_dv)
    access_controls_dv.add('F6:F155')

    distribution_dv = create_data_validation(['Yes', 'Partial', 'No', 'N/A'])
    ws.add_data_validation(distribution_dv)
    distribution_dv.add('G6:G155')

    marking_dv = create_data_validation(['Yes', 'No', 'Inconsistent'])
    ws.add_data_validation(marking_dv)
    marking_dv.add('H6:H155')

    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Sheet protection removed (SRC-018)


def create_communication_tracking_sheet(wb, styles):
    """
    Sheet 6: Communication Tracking
    
    Policy publication, acknowledgment, and training integration.
    """
    ws = wb.create_sheet("Communication Tracking")
    ws.sheet_view.showGridLines = False
    
    # --- Header ---
    merge_and_style(ws, 'A1:P1', 'COMMUNICATION TRACKING', styles['header_main'])
    merge_and_style(ws, 'A2:P2',
                   'Policy publication, acknowledgment, and training integration verification',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35
    
    # --- Column Headers ---
    headers = [
        ('A', 'Policy ID', 20),
        ('B', 'Policy Title', 35),
        ('C', 'Acknowledgment Required', 15),
        ('D', 'Target Audience', 20),
        ('E', 'Publication Date', 12),
        ('F', 'Publication Method', 30),
        ('G', 'Accessibility Verified', 15),
        ('H', 'Acknowledgment Mechanism', 20),
        ('I', 'Acknowledgment Rate', 12),
        ('J', 'Acknowledgment Timeframe', 15),
        ('K', 'Non Acknowledgment Follow Up', 20),
        ('L', 'Training Integration', 15),
        ('M', 'User Feedback Mechanism', 15),
        ('N', 'Communication Compliance Rating', 20),
        ('O', 'Gap Description', 40),
        ('P', 'Evidence Reference', 30)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)

    # --- Sample Row (Row 5, F2F2F2 grey) ---
    _smp_fill = PatternFill('solid', fgColor='F2F2F2')
    _smp_font = Font(name='Calibri', size=10, italic=True, color='808080')
    _smp_bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
    _ct_sample = {
        'A': 'POL-EXAMPLE', 'B': 'Information Security Policy (Example)',
        'C': 'Yes', 'D': 'All Staff',
        'E': '01.01.2025', 'F': 'SharePoint + Email Announcement',
        'G': 'Yes', 'H': 'LMS', 'I': 0.92,
        'J': '30-Days', 'K': 'Yes', 'L': 'Yes', 'M': 'Yes',
        'N': 'Compliant', 'O': '', 'P': 'Example — replace with actual communication data',
    }
    for col_letter in [c[0] for c in headers]:
        ws[f'{col_letter}5'].value = _ct_sample.get(col_letter, '')
        ws[f'{col_letter}5'].fill = _smp_fill
        ws[f'{col_letter}5'].font = _smp_font
        ws[f'{col_letter}5'].border = _smp_bdr
    ws['I5'].number_format = '0%'

    # --- Data Rows (6-155) ---
    for row in range(6, 156):
        # Auto-populate columns
        ws[f'A{row}'] = f'=IF(ROW()-5<=COUNTA(\'Policy Inventory\'!$A:$A)-4,INDEX(\'Policy Inventory\'!$A:$A,ROW()-1),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},\'Policy Inventory\'!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])

        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},\'Policy Inventory\'!$A:$O,15,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])

        ws[f'I{row}'] = f'=IF(A{row}="","",VLOOKUP(A{row},\'Policy Inventory\'!$A:$P,16,FALSE))'
        apply_cell_style(ws[f'I{row}'], styles['calculated_cell'])
        ws[f'I{row}'].number_format = '0%'

        # Manual entry cells
        for col in ['D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'M', 'O', 'P']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # Communication_Compliance_Rating (N) - auto-calculated
        ws[f'N{row}'] = (f'=IF(B{row}="","",IF(AND(G{row}="Yes",OR(C{row}="No",AND(C{row}<>"No",I{row}>=0.9)),L{row}<>"No"),"Compliant",'
                        f'IF(OR(G{row}="No",AND(C{row}<>"No",I{row}<0.7)),"Non-Compliant","Partial")))')
        apply_cell_style(ws[f'N{row}'], styles['calculated_cell'])

    # Date formatting
    for row in range(6, 156):
        ws[f'E{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    accessibility_dv = create_data_validation(['Yes', 'Partial', 'No'])
    ws.add_data_validation(accessibility_dv)
    accessibility_dv.add('G6:G155')

    ack_mech_dv = create_data_validation(['N/A', 'LMS', 'Email', 'E-Signature', 'Form', 'Training'])
    ws.add_data_validation(ack_mech_dv)
    ack_mech_dv.add('H6:H155')

    ack_time_dv = create_data_validation(['N/A', 'Immediate', '30-Days', '60-Days', '90-Days', 'Annual'])
    ws.add_data_validation(ack_time_dv)
    ack_time_dv.add('J6:J155')

    followup_dv = create_data_validation(['Yes', 'Partial', 'No', 'N/A'])
    ws.add_data_validation(followup_dv)
    followup_dv.add('K6:K155')

    training_dv = create_data_validation(['Yes', 'Partial', 'No', 'N/A'])
    ws.add_data_validation(training_dv)
    training_dv.add('L6:L155')

    feedback_dv = create_data_validation(['Yes', 'No'])
    ws.add_data_validation(feedback_dv)
    feedback_dv.add('M6:M55')

    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Sheet protection removed (SRC-018)


def create_repository_assessment_sheet(wb, styles):
    """
    Sheet 7: Repository Assessment
    
    Policy repository structure, organisation, and accessibility (repository-wide).
    """
    ws = wb.create_sheet("Repository Assessment")
    ws.sheet_view.showGridLines = False
    
    # --- Header ---
    merge_and_style(ws, 'A1:B1', 'REPOSITORY ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:B2',
                   'Policy repository structure, organisation, access, and performance',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35
    
    # --- Assessment Questions (Rows 5-22) ---
    questions = [
        (5, 'Repository_Type', 'What system is used for policy repository?'),
        (6, 'Repository_URL_Path', 'What is the primary repository location?'),
        (7, 'Repository_Organisation', 'How is repository organised?'),
        (8, 'Navigation_Ease', 'How easy is it to find policies?'),
        (9, 'Search_Functionality', 'Is search functionality available?'),
        (10, 'Version_Control', 'How are policy versions managed?'),
        (11, 'Previous_Versions_Accessible', 'Can users access previous versions?'),
        (12, 'Access_Control_Implementation', 'Are access controls implemented?'),
        (13, 'Access_Logging', 'Is policy access logged?'),
        (14, 'Mobile_Accessibility', 'Are policies accessible from mobile?'),
        (15, 'Offline_Access', 'Can policies be accessed offline?'),
        (16, 'Repository_Backup', 'Is repository backed up?'),
        (17, 'Repository_Disaster_Recovery', 'Is repository in DR plan?'),
        (18, 'Repository_Performance', 'Is repository performance acceptable?'),
        (19, 'Repository_Uptime', 'What is repository availability?'),
        (20, 'Repository_Compliance_Rating', 'Overall repository compliance?'),
        (21, 'Gap_Description', 'Describe any repository gaps'),
        (22, 'Evidence_Reference', 'Where is repository evidence?')
    ]
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 60
    
    for row_num, field_name, question in questions:
        ws[f'A{row_num}'] = question
        apply_cell_style(ws[f'A{row_num}'], styles['label_cell'])
        
        if row_num == 20:  # Compliance Rating - auto-calculated
            ws[f'B{row_num}'] = ('=IF(AND(B7<>"Unorganised",B8<>"Poor",B9<>"None",B10<>"No-Versioning",'
                                'B12="Yes",B16="Yes"),"Compliant",IF(OR(B7="Unorganised",B8="Poor",'
                                'B10="No-Versioning",B12="No",B16="No"),"Non-Compliant","Partial"))')
            apply_cell_style(ws[f'B{row_num}'], styles['calculated_cell'])
        else:
            apply_cell_style(ws[f'B{row_num}'], styles['input_cell'])
    
    # --- Data Validation ---
    # Repository_Type (B5)
    repo_type_dv = create_data_validation(['SharePoint', 'Confluence', 'Document-Management-System', 
                                          'Intranet', 'File-Share', 'Other'])
    ws.add_data_validation(repo_type_dv)
    repo_type_dv.add('B5')
    
    # Repository_Organisation (B7)
    repo_org_dv = create_data_validation(['By-Hierarchy-Tier', 'By-Domain', 'By-ISO-Control', 
                                         'Alphabetical', 'Unorganised', 'Mixed'])
    ws.add_data_validation(repo_org_dv)
    repo_org_dv.add('B7')
    
    # Navigation_Ease (B8)
    nav_dv = create_data_validation(['Excellent', 'Good', 'Adequate', 'Poor'])
    ws.add_data_validation(nav_dv)
    nav_dv.add('B8')
    
    # Search_Functionality (B9)
    search_dv = create_data_validation(['Excellent', 'Good', 'Adequate', 'None'])
    ws.add_data_validation(search_dv)
    search_dv.add('B9')
    
    # Version_Control (B10)
    version_dv = create_data_validation(['Automated-Versioning', 'Manual-Versioning', 'No-Versioning'])
    ws.add_data_validation(version_dv)
    version_dv.add('B10')
    
    # Previous_Versions_Accessible (B11)
    prev_ver_dv = create_data_validation(['Yes-All-Versions', 'Yes-Recent-Versions', 'No', 'N/A'])
    ws.add_data_validation(prev_ver_dv)
    prev_ver_dv.add('B11')
    
    # Yes/Partial/No dropdowns
    ypn_dv = create_data_validation(['Yes', 'Partial', 'No'])
    ws.add_data_validation(ypn_dv)
    for row in [12, 13]:
        ypn_dv.add(f'B{row}')
    
    # Mobile_Accessibility (B14)
    mobile_dv = create_data_validation(['Yes-Fully', 'Yes-Partially', 'No', 'Not-Tested'])
    ws.add_data_validation(mobile_dv)
    mobile_dv.add('B14')
    
    # Offline_Access (B15)
    offline_dv = create_data_validation(['Yes', 'No', 'Not-Required'])
    ws.add_data_validation(offline_dv)
    offline_dv.add('B15')
    
    # Yes/No/Unknown dropdowns
    ynu_dv = create_data_validation(['Yes', 'No', 'Unknown'])
    ws.add_data_validation(ynu_dv)
    for row in [16, 17]:
        ynu_dv.add(f'B{row}')
    
    # Repository_Performance (B18)
    perf_dv = create_data_validation(['Excellent', 'Good', 'Adequate', 'Poor'])
    ws.add_data_validation(perf_dv)
    perf_dv.add('B18')
    
    # Repository_Uptime (B19)
    uptime_dv = create_data_validation(['99.9%+', '99-99.9%', '95-99%', '<95%', 'Unknown'])
    ws.add_data_validation(uptime_dv)
    uptime_dv.add('B19')
    
    # Sheet protection removed (SRC-018)


def create_gap_analysis_sheet(wb, styles):
    """
    Sheet 8: Gap Analysis
    
    Consolidated gaps from all assessment domains.
    """
    ws = wb.create_sheet("Gap Analysis")
    ws.sheet_view.showGridLines = False
    
    # --- Header ---
    merge_and_style(ws, 'A1:P1', 'GAP ANALYSIS', styles['header_main'])
    merge_and_style(ws, 'A2:P2',
                   'Consolidated gaps from all assessment domains with risk levels and remediation plans',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35
    
    # --- Column Headers ---
    headers = [
        ('A', 'Gap ID', 12),
        ('B', 'Policy ID', 20),
        ('C', 'Policy Title', 35),
        ('D', 'Gap Category', 15),
        ('E', 'Gap Description', 40),
        ('F', 'Risk Level', 15),
        ('G', 'Impact Assessment', 35),
        ('H', 'Affected Stakeholders', 25),
        ('I', 'Remediation Action', 40),
        ('J', 'Responsible Party', 25),
        ('K', 'Target Completion Date', 15),
        ('L', 'Estimated Effort', 15),
        ('M', 'Dependencies', 30),
        ('N', 'Status', 15),
        ('O', 'Completion Evidence', 30),
        ('P', 'Risk Acceptance', 40)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)

    # --- Sample Row (Row 5, F2F2F2 grey) ---
    _smp_fill = PatternFill('solid', fgColor='F2F2F2')
    _smp_font = Font(name='Calibri', size=10, italic=True, color='808080')
    _smp_bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
    _gap_sample = {
        'A': 'GAP-EXAMPLE', 'B': 'POL-001', 'C': 'Information Security Policy (Example)',
        'D': 'Lifecycle', 'E': 'Policy has not been reviewed in 18 months (example)',
        'F': 'High', 'G': 'Regulatory non-compliance risk (example)',
        'H': 'CISO, Legal', 'I': 'Schedule annual review and update policy (example)',
        'J': 'Policy Owner', 'K': '30.06.2026', 'L': '1day',
        'M': 'None', 'N': 'Not-Started', 'O': '', 'P': 'Example — replace with actual gap data',
    }
    for col_letter in [c[0] for c in headers]:
        ws[f'{col_letter}5'].value = _gap_sample.get(col_letter, '')
        ws[f'{col_letter}5'].fill = _smp_fill
        ws[f'{col_letter}5'].font = _smp_font
        ws[f'{col_letter}5'].border = _smp_bdr

    # --- Data Rows (6-105, 100 rows for gaps) ---
    for row in range(6, 106):
        # Gap_ID auto-generated (anchor on col B to suppress phantom values)
        ws[f'A{row}'] = f'=IF(B{row}="","",TEXT(ROW()-5,"GAP-000"))'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # All other columns manual entry
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

    # Date formatting for Target_Completion_Date
    for row in range(6, 106):
        ws[f'K{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    risk_level_dv = create_data_validation(['Critical', 'High', 'Medium', 'Low'])
    ws.add_data_validation(risk_level_dv)
    risk_level_dv.add('F6:F105')

    effort_dv = create_data_validation(['<1hr', '1-4hrs', '1day', '2-5days', '>1week'])
    ws.add_data_validation(effort_dv)
    effort_dv.add('L6:L105')

    status_dv = create_data_validation(['Not-Started', 'In-Progress', 'Blocked', 'Completed', 'Accepted-Risk'])
    ws.add_data_validation(status_dv)
    status_dv.add('N6:N105')

    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Sheet protection removed (SRC-018)


def create_evidence_register(wb):
    """
    Sheet 9: Evidence Register (Gold Standard)
    """
    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_hdr = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _input = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = _border

    columns = [
        ("Evidence ID", 14), ("Evidence Type", 22), ("Description", 45),
        ("Related Control / Section", 28), ("Collection Date (DD.MM.YYYY)", 22),
        ("Storage Location / Reference", 38), ("Collected By", 22), ("Status", 14),
    ]
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    sample_data = [
        "EV-001", "Document", "Sample evidence entry — replace with actual evidence",
        "All Controls", "01.01.2026", "SharePoint/ISMS/Evidence/", "ISMS Team", "Active"
    ]
    for col_idx, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
        cell.fill = _grey_sample
        cell.border = _border

    dv_status = DataValidation(
        type="list",
        formula1='"Active,Archived,Superseded,Pending Review"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(6, 106):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = _input
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


def create_action_items_sheet(wb, styles):
    """
    Sheet 10: Action Items
    
    Remediation tracking (auto-populated from Gap Analysis).
    """
    ws = wb.create_sheet("Action Items")
    ws.sheet_view.showGridLines = False
    
    # --- Header ---
    merge_and_style(ws, 'A1:M1', 'ACTION ITEMS', styles['header_main'])
    merge_and_style(ws, 'A2:M2', 'Remediation tracking for identified gaps', styles['header_sub'])
    
    # --- Column Headers ---
    headers = [
        ('A', 'Action ID', 12),
        ('B', 'Related Gap ID', 12),
        ('C', 'Action Description', 40),
        ('D', 'Priority', 15),
        ('E', 'Owner', 25),
        ('F', 'Target Date', 15),
        ('G', 'Status', 15),
        ('H', 'Progress %', 12),
        ('I', 'Last Update Date', 12),
        ('J', 'Last Update Notes', 35),
        ('K', 'Blocking Issues', 35),
        ('L', 'Escalation Required', 15),
        ('M', 'Escalation To', 25)
    ]
    
    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width
    
    set_column_widths(ws, widths_dict)

    # --- Sample Row (Row 5, F2F2F2 grey) ---
    _smp_fill = PatternFill('solid', fgColor='F2F2F2')
    _smp_font = Font(name='Calibri', size=10, italic=True, color='808080')
    _smp_bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
    _ai_sample = {
        'A': 'ACT-EXAMPLE', 'B': 'GAP-EXAMPLE',
        'C': 'Schedule annual review and update policy (example)',
        'D': 'High', 'E': 'Policy Owner', 'F': '30.06.2026',
        'G': 'Not-Started', 'H': 0.0, 'I': '01.01.2026',
        'J': 'Action created (example)', 'K': 'None', 'L': 'No', 'M': '',
    }
    for col_letter in [c[0] for c in headers]:
        ws[f'{col_letter}5'].value = _ai_sample.get(col_letter, '')
        ws[f'{col_letter}5'].fill = _smp_fill
        ws[f'{col_letter}5'].font = _smp_font
        ws[f'{col_letter}5'].border = _smp_bdr
    ws['H5'].number_format = '0%'
    ws['F5'].number_format = 'DD.MM.YYYY'
    ws['I5'].number_format = 'DD.MM.YYYY'

    # --- Data Rows (6-55, 50 action items) ---
    for row in range(6, 56):
        # Auto-populate from Gap Analysis
        ws[f'A{row}'] = f'=IF(\'Gap Analysis\'!A{row}="","",SUBSTITUTE(\'Gap Analysis\'!A{row},"GAP","ACT"))'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        ws[f'B{row}'] = f'=\'Gap Analysis\'!A{row}'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])

        ws[f'C{row}'] = f'=IF(\'Gap Analysis\'!I{row}<>"",\'Gap Analysis\'!I{row},"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])

        ws[f'D{row}'] = f'=IF(\'Gap Analysis\'!F{row}<>"",\'Gap Analysis\'!F{row},"")'
        apply_cell_style(ws[f'D{row}'], styles['calculated_cell'])

        ws[f'E{row}'] = f'=IF(\'Gap Analysis\'!J{row}<>"",\'Gap Analysis\'!J{row},"")'
        apply_cell_style(ws[f'E{row}'], styles['calculated_cell'])

        ws[f'F{row}'] = f'=IF(\'Gap Analysis\'!K{row}<>"",\'Gap Analysis\'!K{row},"")'
        apply_cell_style(ws[f'F{row}'], styles['calculated_cell'])
        ws[f'F{row}'].number_format = 'DD.MM.YYYY'

        ws[f'G{row}'] = f'=IF(\'Gap Analysis\'!N{row}<>"",\'Gap Analysis\'!N{row},"")'
        apply_cell_style(ws[f'G{row}'], styles['calculated_cell'])

        # Manual entry cells
        for col in ['H', 'I', 'J', 'K', 'L', 'M']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

    # Percentage formatting for Progress %
    for row in range(6, 56):
        ws[f'H{row}'].number_format = '0%'
        ws[f'I{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    escalation_dv = create_data_validation(['Yes', 'No'])
    ws.add_data_validation(escalation_dv)
    escalation_dv.add('L6:L55')

    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Sheet protection removed (SRC-018)


def create_approval_sheet(wb):
    """Create the Approval Sign-Off sheet — Gold Standard."""
    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Summary fields (rows 4-8)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G7:G11),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Assessment Status dropdown (row 7)
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # 3 Approver sections (start at row 11)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ==============================================================================
# VALIDATION FINALISATION
# ==============================================================================

def create_summary_dashboard_sheet(wb):
    """Add Summary Dashboard sheet (TABLE 1/2/3) to workbook."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title="Summary Dashboard")
    ws.sheet_view.showGridLines = False

    def _f(hex_c):
        return PatternFill("solid", fgColor=hex_c)

    def _b():
        s = Side(style="thin", color="000000")
        return Border(left=s, right=s, top=s, bottom=s)

    def _hdr(row, col, value, fill="003366", fc="FFFFFF", sz=11,
             bold=True, merge_to=None):
        c = ws.cell(row=row, column=col)
        c.value = value
        c.font = Font(name="Calibri", bold=bold, color=fc, size=sz)
        c.fill = _f(fill)
        c.alignment = Alignment(horizontal="left" if merge_to else "center",
                                vertical="center", wrap_text=True)
        c.border = _b()
        if merge_to:
            ws.merge_cells(f"{get_column_letter(col)}{row}:{merge_to}")
        return c

    def _dat(row, col, value, fill="FFFFCC", fc="000000", bold=False,
             num=False):
        c = ws.cell(row=row, column=col)
        c.value = value
        c.font = Font(name="Calibri", bold=bold, color=fc)
        c.fill = _f(fill)
        c.alignment = Alignment(
            horizontal="center" if num else "left",
            vertical="center", wrap_text=True)
        c.border = _b()
        return c

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "A3"

    # Row 1: Title (GS-SD-014: must contain "— SUMMARY DASHBOARD" em dash)
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "POLICY FRAMEWORK ASSESSMENT \u2014 SUMMARY DASHBOARD"
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    c.fill = _f("003366")
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    c.border = _b()
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle (gold standard: no fill, left-aligned, color 003366)
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = ("ISO/IEC 27001:2022 | Control A.5.1-2 | "
               "Policy framework compliance status and governance overview")
    c.font = Font(name="Calibri", italic=True, color="003366", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _b()

    # ── TABLE 1: POLICY STATUS DISTRIBUTION ──────────────────────────────
    _hdr(4, 1, "TABLE 1: POLICY STATUS DISTRIBUTION", "003366",
         sz=12, merge_to="F4")
    _hdr(5, 1, "Policy Inventory — Column M (Policy Status), rows 5:54",
         "4472C4", sz=10, merge_to="F5")
    _hdr(6, 1, "Status", "4472C4", sz=10)
    _hdr(6, 2, "Count", "4472C4", sz=10)
    _hdr(6, 3, "% of Total", "4472C4", sz=10)

    t1_statuses = ["Active", "Draft", "Under-Review", "Retired", "Superseded"]
    for i, status in enumerate(t1_statuses):
        r = 7 + i
        _dat(r, 1, status)
        _dat(r, 2, f'=COUNTIF(\'Policy Inventory\'!M5:M54,\"{status}\")',
             num=True)
        _dat(r, 3, f'=IF($B$12=0,"—",TEXT(B{r}/$B$12,"0.0%"))', num=True)

    _dat(12, 1, "TOTAL", "D9D9D9", bold=True)
    _dat(12, 2, "=SUM(B7:B11)", "D9D9D9", num=True)
    _dat(12, 3, "—", "D9D9D9", num=True)

    # ── TABLE 2: KEY METRICS ──────────────────────────────────────────────
    _hdr(14, 1, "TABLE 2: KEY METRICS", "003366", sz=12, merge_to="F14")
    _hdr(15, 1, "KPI Metric", "D9D9D9", fc="000000", sz=10)
    _hdr(15, 2, "Value", "D9D9D9", fc="000000", sz=10)
    _hdr(15, 3, "Notes", "D9D9D9", fc="000000", sz=10, merge_to="F15")

    kpis = [
        ("Total Policies Registered",
         "=COUNTA(\'Policy Inventory\'!A5:A54)",
         "All policies catalogued in Policy Inventory"),
        ("Active Policies",
         "=COUNTIF(\'Policy Inventory\'!M5:M54,\"Active\")",
         "Policies with Active status"),
        ("Policies Under Review",
         "=COUNTIF(\'Policy Inventory\'!M5:M54,\"Under-Review\")",
         "Policies currently under review — confirm progress"),
        ("Draft Policies",
         "=COUNTIF(\'Policy Inventory\'!M5:M54,\"Draft\")",
         "Draft policies not yet published or approved"),
        ("Critical Gaps Identified",
         "=COUNTIF(\'Gap Analysis\'!F6:F105,\"Critical\")",
         "Critical-risk gaps in Gap Analysis sheet"),
        ("Open Remediation Actions",
         "=COUNTIF(\'Gap Analysis\'!N6:N105,\"Not-Started\")"
         "+COUNTIF(\'Gap Analysis\'!N6:N105,\"In-Progress\")",
         "Gaps not yet resolved (Not-Started + In-Progress)"),
    ]
    for i, (metric, formula, note) in enumerate(kpis):
        r = 16 + i
        _dat(r, 1, metric)
        _dat(r, 2, formula, num=True)
        c = ws.cell(row=r, column=3)
        c.value = note
        c.font = Font(name="Calibri", color="000000")
        c.fill = _f("FFFFCC")
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        c.border = _b()
        ws.merge_cells(f"C{r}:F{r}")

    # ── TABLE 3: CRITICAL FINDINGS ────────────────────────────────────────
    _hdr(23, 1, "TABLE 3: CRITICAL FINDINGS", "C00000", sz=12, merge_to="F23")
    _hdr(24, 1, "Finding", "D9D9D9", fc="000000", sz=10)
    _hdr(24, 2, "Count", "D9D9D9", fc="000000", sz=10)
    _hdr(24, 3, "Action Required", "D9D9D9", fc="000000", sz=10,
         merge_to="F24")

    critical = [
        ("Policies with Critical Gap",
         "=COUNTIF(\'Policy Inventory\'!S5:S54,\"Critical-Gap\")",
         "Review and remediate policies flagged as Critical-Gap in column S"),
        ("Retired or Superseded Policies",
         "=COUNTIF(\'Policy Inventory\'!M5:M54,\"Retired\")"
         "+COUNTIF(\'Policy Inventory\'!M5:M54,\"Superseded\")",
         "Confirm retired/superseded policies are archived and no longer in use"),
        ("Policies Under Review (overdue check)",
         "=COUNTIF(\'Policy Inventory\'!M5:M54,\"Under-Review\")",
         "Confirm all Under-Review policies have an assigned owner and target date"),
    ]
    for i, (finding, formula, action) in enumerate(critical):
        r = 25 + i
        c = ws.cell(row=r, column=1)
        c.value = finding
        c.font = Font(name="Calibri", bold=True, color="000000")
        c.fill = _f("FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        c.border = _b()
        c2 = ws.cell(row=r, column=2)
        c2.value = formula
        c2.font = Font(name="Calibri", color="000000")
        c2.fill = _f("FFFFFF")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = _b()
        c3 = ws.cell(row=r, column=3)
        c3.value = action
        c3.font = Font(name="Calibri", color="000000")
        c3.fill = _f("FFFFFF")
        c3.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True)
        c3.border = _b()
        ws.merge_cells(f"C{r}:F{r}")



def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly


# ==============================================================================
# MAIN EXECUTION
# ==============================================================================



def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Review the Dashboard (auto-calculated) for a high-level compliance overview.',
        '2. Complete the Policy Inventory — list all information security policies with ownership and review dates.',
        '3. Complete Lifecycle Compliance — verify approval, review cycle, and publication status per policy.',
        '4. Complete Governance Assessment — confirm ownership, RACI accountability, and authority structures.',
        '5. Complete Classification Review — verify classification levels and access control alignment.',
        '6. Complete Communication Tracking — record policy acknowledgement rates and distribution evidence.',
        '7. Complete Repository Assessment — verify policy repository accessibility and version control.',
        '8. Review the Gap Analysis (auto-populated) — prioritise remediation by risk level.',
        '9. Attach audit evidence in the Evidence Register (EV-xxx references).',
        '10. Create remediation actions in Action Items with owners and target dates.',
        '11. Obtain three-level stakeholder approval in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A25"] = "Status Legend"
    ws["A25"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=26, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 27 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.5.1-2-6.1-2.S1 - Policy Framework Assessment")
    logger.info("Workbook Generator")
    logger.info("=" * 80)

    # Initialize workbook
    logger.info("Creating workbook...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default blank sheet (Dashboard removed — Summary Dashboard serves this purpose)
    styles = _STYLES

    # Create all sheets

    logger.info("Creating Instructions & Legend...")
    create_instructions_sheet(wb.create_sheet())
    logger.info("Creating Policy Inventory...")
    create_policy_inventory_sheet(wb, styles)

    logger.info("Creating Lifecycle Compliance...")
    create_lifecycle_compliance_sheet(wb, styles)

    logger.info("Creating Governance Assessment...")
    create_governance_assessment_sheet(wb, styles)

    logger.info("Creating Classification Review...")
    create_classification_review_sheet(wb, styles)

    logger.info("Creating Communication Tracking...")
    create_communication_tracking_sheet(wb, styles)

    logger.info("Creating Repository Assessment...")
    create_repository_assessment_sheet(wb, styles)

    logger.info("Creating Gap Analysis...")
    create_gap_analysis_sheet(wb, styles)

    logger.info("Creating Evidence Register...")
    create_evidence_register(wb)

    logger.info("Creating Action Items...")
    create_action_items_sheet(wb, styles)

    logger.info("Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb)

    logger.info("Creating Approval Sign-Off...")
    create_approval_sheet(wb)

    # Set workbook properties
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.keywords = "Policy, Governance, ISMS, ISO27001, A.5.1"
    wb.properties.comments = "Generated via Python script generate_a5_1_2_6_1_2_s1_policy_framework.py"

    # Finalise data validations
    finalize_validations(wb)

    # Generate filename with current date
    today = datetime.now().strftime("%Y%m%d")

    # Save workbook to WKBK directory
    logger.info(f"Saving workbook as: {output_path.name}")
    wb.save(output_path)
    logger.info("=" * 80)
    logger.info("Workbook generated successfully!")
    logger.info("=" * 80)
    logger.info("Next Steps:")
    logger.info("1. Open the workbook in Excel")
    logger.info("2. Complete Sheet 2 (Policy Inventory) - this is your foundation")
    logger.info("3. Complete Sheets 3-7 (assessment domains)")
    logger.info("4. Review Sheet 8 (Gap Analysis) - auto-populated")
    logger.info("5. Document Sheet 9 (Evidence)")
    logger.info("6. Plan Sheet 10 (Action Items)")
    logger.info("7. Review Sheet 1 (Dashboard) - auto-calculated")
    logger.info("8. Obtain Sheet 11 (Approval Sign-Off)")
    logger.info(f"File location: ./{output_path.name}")
def main():
    try:
        create_workbook(_wkbk_dir / OUTPUT_FILENAME)
    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================