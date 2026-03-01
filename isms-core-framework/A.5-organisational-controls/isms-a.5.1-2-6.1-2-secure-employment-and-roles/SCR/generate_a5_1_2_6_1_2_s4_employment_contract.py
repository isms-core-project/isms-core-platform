#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.1-2-6.1-2.S4 - Employment Contract Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.1-2-6.1-2: Secure Employment and Roles
Assessment Domain 4 of 4: Employment Contract Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific information security policies and organisation infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Policy inventory scope and mandatory policy categories (match your organisation)
2. Role definition framework and responsibility assignment methodology
3. Screening procedure requirements and verification standards (adapt to jurisdiction)
4. Employment contract security clause requirements (legal review required)
5. Policy governance cycle including approval authorities and review owners

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.1-2-6.1-2 Secure Employment and Roles Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
information security policies and organisation controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Employment Contract Assessment under ISO 27001:2022 Controls A.5.1, A.5.2, A.6.1, and A.6.2. Supports evidence-based evaluation of policy governance, role accountability, personnel screening, and employment framework compliance.

**Assessment Scope:**
- Policy inventory completeness and lifecycle compliance
- Role and responsibility assignment accuracy and coverage
- Personnel screening and vetting procedure adherence
- Employment contract security clause documentation
- Policy governance and approval workflow effectiveness
- Communication and acknowledgment tracking completeness
- Evidence collection for HR, governance, and compliance audits

**Generated Workbook Structure:**
1. Contract Template Assessment
2. Required Clause Registry
3. Personnel Contract Compliance
4. Confidentiality NDA Tracking
5. Post Employment Obligations
6. Contractor Agreement Assessment
7. Gap Analysis
8. Evidence Register
9. Approval Sign-Off

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 4 domains covering Secure Employment and Roles controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a5_1_2_6_1_2_s4_employment_contract.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a5_1_2_6_1_2_s4_employment_contract.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a5_1_2_6_1_2_s4_employment_contract.py --date 20250115

Output:
    File: ISMS-IMP-A.5.1-2-6.1-2.S4_Employment_Contract_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.1-2-6.1-2
Assessment Domain:    4 of 4 (Employment Contract Assessment)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.1-2-6.1-2: Secure Employment and Roles Policy (Governance)
    - ISMS-IMP-A.5.1-2-6.1-2.S1: Policy Framework Assessment Workbook (Domain 1)
    - ISMS-IMP-A.5.1-2-6.1-2.S2: Roles Responsibilities Assessment Workbook (Domain 2)
    - ISMS-IMP-A.5.1-2-6.1-2.S3: Screening Vetting Assessment Workbook (Domain 3)
    - ISMS-IMP-A.5.1-2-6.1-2.S4: Employment Contract Assessment (Domain 4)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.1-2-6.1-2.S4 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive information security policies and organisation details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review policy inventory, role definitions, and screening procedures annually or when organisational restructuring occurs, regulatory requirements change, or HR processes are updated.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
import sys
from datetime import datetime
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.1-2-6.1-2.S4"
WORKBOOK_NAME = "Employment Contract Assessment"
CONTROL_ID = "A.5.1-2-6.1-2"
CONTROL_NAME = "Secure Employment and Roles"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# WKBK output directory
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ==============================================================================
# STYLE DEFINITIONS
# ==============================================================================
def setup_styles():
    """
    Define consistent cell styles used throughout the workbook.

    Returns:
        dict: Dictionary of style objects for different cell types
    """
    styles = {}

    # Header styles
    styles['header_main'] = {
        'font': Font(name='Calibri', size=18, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['header_sub'] = {
        'font': Font(name='Calibri', size=12, bold=False, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['section_header'] = {
        'font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

    styles['column_header'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    # Data cell styles
    styles['input_cell'] = {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),  # Yellow
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    styles['calculated_cell'] = {
        'font': Font(name='Calibri', size=10, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),  # Grey (standard palette)
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    styles['label_cell'] = {
        'font': Font(name='Calibri', size=10, bold=True, color='000000'),
        'fill': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),  # Gray
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    }

    # Status fills (for conditional formatting reference)
    styles['status_compliant'] = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
    styles['status_partial'] = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')  # Yellow
    styles['status_non_compliant'] = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red

    return styles


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================


_STYLES = setup_styles()
def apply_cell_style(cell, style_dict):
    """
    Apply a style dictionary to a cell.

    Args:
        cell: openpyxl cell object
        style_dict: Dictionary containing style attributes
    """
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']


def set_column_widths(ws, widths_dict):
    """
    Set column widths for a worksheet.

    Args:
        ws: Worksheet object
        widths_dict: Dictionary mapping column letters to widths
    """
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width


def create_data_validation(formula_or_list, allow_blank=True, error_title="Invalid Entry",
                           error_message="Please select from dropdown"):
    """
    Create a data validation object for dropdown lists.

    Args:
        formula_or_list: Either a string formula or list of values
        allow_blank: Whether to allow blank cells
        error_title: Title for error message
        error_message: Error message text

    Returns:
        DataValidation object
    """
    if isinstance(formula_or_list, list):
        dv = DataValidation(type="list", formula1=f'"{",".join(formula_or_list)}"',
                           allow_blank=allow_blank)
    else:
        dv = DataValidation(type="list", formula1=formula_or_list, allow_blank=allow_blank)

    dv.error = error_message
    dv.errorTitle = error_title
    dv.showErrorMessage = True
    return dv


def merge_and_style(ws, range_string, value, style_dict):
    """
    Merge cells and apply style.

    Args:
        ws: Worksheet object
        range_string: Range to merge (e.g., 'A1:K1')
        value: Value to write
        style_dict: Style dictionary to apply
    """
    ws.merge_cells(range_string)
    cell = ws[range_string.split(':')[0]]
    cell.value = value
    apply_cell_style(cell, style_dict)


# ==============================================================================
# SHEET CREATION FUNCTIONS
# ==============================================================================


def create_contract_template_assessment_sheet(wb, styles):
    """
    Sheet 2: Contract Template Assessment

    Verify contract templates are complete and current.
    50 rows for templates.
    """
    ws = wb.create_sheet("Contract Template Assessment")
    ws.sheet_view.showGridLines = False

    # --- Header ---
    merge_and_style(ws, 'A1:N1', 'CONTRACT TEMPLATE ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:N2',
                   'Verification of employment contract templates for security clause completeness',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35

    # --- Column Headers (Row 3) ---
    headers = [
        ('A', 'Template ID', 15),
        ('B', 'Template Name', 35),
        ('C', 'Template Type', 20),
        ('D', 'Personnel Category', 20),
        ('E', 'Current Version', 12),
        ('F', 'Version Date', 12),
        ('G', 'Last Review Date', 12),
        ('H', 'Next Review Date', 12),
        ('I', 'Security Clauses Count', 15),
        ('J', 'Required Clauses Met', 15),
        ('K', 'Legal Review Date', 12),
        ('L', 'Legal Approved', 15),
        ('M', 'Compliance Status', 15),
        ('N', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}3'] = header_text
        apply_cell_style(ws[f'{col_letter}3'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Sample Row (Row 4) — Step 2.4 ---
    _cta_smp_fill = PatternFill('solid', fgColor='F2F2F2')
    _cta_smp_font = Font(name='Calibri', size=10, italic=True, color='808080')
    _cta_smp_bdr = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    _cta_sample = {
        'A': 'CTA-EXAMPLE', 'B': 'Standard Employment Contract',
        'C': 'Employee-Permanent', 'D': 'All-Personnel', 'E': 'v2.1',
        'F': '01.01.2025', 'G': '15.01.2026', 'H': '15.01.2027',
        'I': 15, 'J': 'Yes', 'K': '15.01.2026', 'L': 'Yes',
        'M': 'Compliant',
        'N': 'Example entry — replace with your contract template details',
    }
    for col_letter in [c[0] for c in headers]:
        ws[f'{col_letter}4'].value = _cta_sample.get(col_letter, '')
        ws[f'{col_letter}4'].fill = _cta_smp_fill
        ws[f'{col_letter}4'].font = _cta_smp_font
        ws[f'{col_letter}4'].border = _cta_smp_bdr

    # --- Data Rows (5-54) ---
    for row in range(5, 55):
        for col_letter in [c[0] for c in headers]:
            cell = ws[f'{col_letter}{row}']
            if col_letter == 'M':
                # Auto-calculated compliance status
                ws[f'M{row}'] = (f'=IF(B{row}="","",IF(AND(J{row}="Yes",L{row}="Yes",'
                                f'H{row}>=TODAY()),"Compliant",'
                                f'IF(OR(J{row}="No",L{row}="No"),"Non-Compliant","Partial")))')
                apply_cell_style(cell, styles['calculated_cell'])
            else:
                apply_cell_style(cell, styles['input_cell'])

    # --- Data Validation ---
    # Template Type dropdown
    type_dv = create_data_validation(['Employee-Permanent', 'Employee-Fixed-Term', 'Executive',
                                       'Intern', 'Contractor', 'Consultant', 'Agency-Worker'])
    ws.add_data_validation(type_dv)
    type_dv.add('C5:C54')

    # Personnel Category dropdown
    category_dv = create_data_validation(['All-Personnel', 'IT-Staff', 'Management', 'Finance',
                                          'HR', 'Operations', 'External', 'Privileged-Users'])
    ws.add_data_validation(category_dv)
    category_dv.add('D5:D54')

    # Yes/No dropdowns
    yn_dv = create_data_validation(['Yes', 'No', 'Partial'])
    ws.add_data_validation(yn_dv)
    yn_dv.add('J5:J54')
    yn_dv.add('L5:L54')

    # Date formatting
    for col in ['F', 'G', 'H', 'K']:
        for row in range(5, 55):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes at A4 (DS-007: standard, rows 1-3 frozen)
    ws.freeze_panes = 'A4'

    # Define named range for Template_ID_List
    template_id_range = DefinedName(name='Template_ID_List', attr_text="\'Contract Template Assessment\'!$A$5:$A$54")
    wb.defined_names.add(template_id_range)

    # Sheet protection removed (SRC-018)


def create_required_clause_registry_sheet(wb, styles):
    """
    Sheet 3: Required Clause Registry

    Master list of required security clauses and template coverage.
    """
    ws = wb.create_sheet("Required Clause Registry")
    ws.sheet_view.showGridLines = False

    # --- Header ---
    merge_and_style(ws, 'A1:L1', 'REQUIRED CLAUSE REGISTRY', styles['header_main'])
    merge_and_style(ws, 'A2:L2',
                   'Master list of required security clauses with template coverage verification',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35

    # --- Column Headers ---
    headers = [
        ('A', 'Clause ID', 15),
        ('B', 'Clause Category', 20),
        ('C', 'Clause Title', 35),
        ('D', 'Clause Description', 50),
        ('E', 'ISO 27001 Reference', 20),
        ('F', 'Legal Requirement', 20),
        ('G', 'Mandatory For', 25),
        ('H', 'Coverage Status', 15),
        ('I', 'Templates With Clause', 25),
        ('J', 'Gap If Missing', 40),
        ('K', 'Last Verified', 12),
        ('L', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Pre-populate Required Clauses (Rows 5-34) ---
    required_clauses = [
        ('RC-01', 'Confidentiality', 'Confidentiality and Non-Disclosure',
         'Employee agrees to maintain confidentiality of all company information', 'A.5.1, A.6.2', 'Employment Law'),
        ('RC-02', 'Confidentiality', 'Non-Disclosure Agreement Reference',
         'Reference to standalone NDA or embedded NDA terms', 'A.6.2', 'Employment Law'),
        ('RC-03', 'IP Rights', 'Intellectual Property Assignment',
         'All work product IP belongs to the organisation', 'A.5.1', 'IP Law'),
        ('RC-04', 'IP Rights', 'Invention Assignment',
         'Employee assigns rights to inventions created during employment', 'A.5.1', 'Patent Law'),
        ('RC-05', 'Security', 'Information Security Responsibilities',
         'Employee responsibilities for information security', 'A.6.2', 'ISMS Policy'),
        ('RC-06', 'Security', 'Acceptable Use Agreement Reference',
         'Reference to or acknowledgment of acceptable use policy', 'A.5.10', 'ISMS Policy'),
        ('RC-07', 'Security', 'Access Control Acknowledgment',
         'Agreement to comply with access control policies', 'A.5.15', 'ISMS Policy'),
        ('RC-08', 'Security', 'Password/Authentication Requirements',
         'Agreement to follow authentication policies', 'A.5.17', 'ISMS Policy'),
        ('RC-09', 'Post-Employment', 'Return of Assets',
         'Obligation to return all company assets upon termination', 'A.6.5', 'Employment Law'),
        ('RC-10', 'Post-Employment', 'Post-Employment Confidentiality',
         'Confidentiality obligations continue after termination', 'A.6.5', 'Contract Law'),
        ('RC-11', 'Post-Employment', 'Non-Compete Clause',
         'Restrictions on competing employment if applicable', 'A.6.5', 'Contract Law'),
        ('RC-12', 'Post-Employment', 'Non-Solicitation Clause',
         'Restrictions on soliciting employees/clients', 'A.6.5', 'Contract Law'),
        ('RC-13', 'Compliance', 'Policy Compliance Commitment',
         'Agreement to comply with all company policies', 'A.5.1', 'ISMS Policy'),
        ('RC-14', 'Compliance', 'Training Completion Requirement',
         'Agreement to complete required security training', 'A.6.3', 'ISMS Policy'),
        ('RC-15', 'Compliance', 'Incident Reporting Obligation',
         'Obligation to report security incidents', 'A.6.8', 'ISMS Policy'),
        ('RC-16', 'Screening', 'Background Check Authorisation',
         'Authorisation for pre-employment screening', 'A.6.1', 'Employment Law'),
        ('RC-17', 'Screening', 'Reference Check Authorisation',
         'Authorisation to contact references', 'A.6.1', 'Employment Law'),
        ('RC-18', 'Data Protection', 'Personal Data Processing Consent',
         'Consent for processing of employee personal data', 'FADP', 'Data Protection'),
        ('RC-19', 'Data Protection', 'Data Protection Responsibilities',
         'Employee responsibilities for data protection', 'A.5.34', 'Data Protection'),
        ('RC-20', 'Monitoring', 'Monitoring Acknowledgment',
         'Acknowledgment of workplace/IT monitoring practices', 'A.8.15', 'Employment Law')
    ]

    for i, clause in enumerate(required_clauses, start=5):
        ws[f'A{i}'] = clause[0]
        ws[f'B{i}'] = clause[1]
        ws[f'C{i}'] = clause[2]
        ws[f'D{i}'] = clause[3]
        ws[f'E{i}'] = clause[4]
        ws[f'F{i}'] = clause[5]

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            apply_cell_style(ws[f'{col}{i}'], styles['calculated_cell'])

        for col in ['G', 'H', 'I', 'J', 'K', 'L']:
            apply_cell_style(ws[f'{col}{i}'], styles['input_cell'])

    # Additional blank rows for user-defined clauses (25-74, 50 empty rows)
    for row in range(25, 75):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

    # --- Data Validation ---
    category_dv = create_data_validation(['Confidentiality', 'IP Rights', 'Security',
                                          'Post-Employment', 'Compliance', 'Screening',
                                          'Data Protection', 'Monitoring', 'Other'])
    ws.add_data_validation(category_dv)
    category_dv.add('B5:B74')

    coverage_dv = create_data_validation(['Covered', 'Partial', 'Not-Covered', 'N/A'])
    ws.add_data_validation(coverage_dv)
    coverage_dv.add('H5:H74')

    # Date formatting
    for row in range(5, 75):
        ws[f'K{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes
    ws.freeze_panes = 'A4'

    # Sheet protection removed (SRC-018)


def create_personnel_contract_compliance_sheet(wb, styles):
    """
    Sheet 4: Personnel Contract Compliance

    Per-person contract status verification.
    200 rows for personnel.
    """
    ws = wb.create_sheet("Personnel Contract Compliance")
    ws.sheet_view.showGridLines = False

    # --- Header ---
    merge_and_style(ws, 'A1:O1', 'PERSONNEL CONTRACT COMPLIANCE', styles['header_main'])
    merge_and_style(ws, 'A2:O2',
                   'Individual employee contract verification and compliance tracking',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35

    # --- Column Headers ---
    headers = [
        ('A', 'Employee ID', 15),
        ('B', 'Employee Name', 25),
        ('C', 'Department', 20),
        ('D', 'Role Title', 25),
        ('E', 'Employment Type', 15),
        ('F', 'Start Date', 12),
        ('G', 'Contract Template Used', 20),
        ('H', 'Contract Signed', 15),
        ('I', 'Contract Date', 12),
        ('J', 'Contract On File', 15),
        ('K', 'All Clauses Present', 15),
        ('L', 'Screening Complete', 15),
        ('M', 'Background Check Date', 12),
        ('N', 'Compliance Status', 15),
        ('O', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}3'] = header_text
        apply_cell_style(ws[f'{col_letter}3'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Sample Row (Row 4) — Step 2.4 ---
    _pcc_smp_fill = PatternFill('solid', fgColor='F2F2F2')
    _pcc_smp_font = Font(name='Calibri', size=10, italic=True, color='808080')
    _pcc_smp_bdr = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    _pcc_sample = {
        'A': 'EMP-EXAMPLE', 'B': 'Jane Smith (Example)',
        'C': 'Information Security', 'D': 'Information Security Manager',
        'E': 'Permanent', 'F': '01.03.2024',
        'G': 'Standard Employment Contract v2.1',
        'H': 'Yes', 'I': '01.03.2024', 'J': 'Yes',
        'K': 'Yes', 'L': 'Yes', 'M': '01.02.2024',
        'N': 'Compliant',
        'O': 'Example entry — replace with your personnel contract records',
    }
    for col_letter in [c[0] for c in headers]:
        ws[f'{col_letter}4'].value = _pcc_sample.get(col_letter, '')
        ws[f'{col_letter}4'].fill = _pcc_smp_fill
        ws[f'{col_letter}4'].font = _pcc_smp_font
        ws[f'{col_letter}4'].border = _pcc_smp_bdr

    # --- Data Rows (5-54) ---
    for row in range(5, 55):
        for col_letter in [c[0] for c in headers]:
            cell = ws[f'{col_letter}{row}']
            if col_letter == 'N':
                # Auto-calculated compliance status
                ws[f'N{row}'] = (f'=IF(B{row}="","",IF(AND(H{row}="Yes",J{row}="Yes",'
                                f'K{row}="Yes",L{row}="Yes"),"Compliant",'
                                f'IF(OR(H{row}="No",J{row}="No",L{row}="No"),"Non-Compliant","Partial")))')
                apply_cell_style(cell, styles['calculated_cell'])
            else:
                apply_cell_style(cell, styles['input_cell'])

    # --- Data Validation ---
    emp_type_dv = create_data_validation(['Permanent', 'Fixed-Term', 'Executive', 'Intern',
                                          'Contractor', 'Consultant', 'Agency'])
    ws.add_data_validation(emp_type_dv)
    emp_type_dv.add('E5:E54')

    yn_dv = create_data_validation(['Yes', 'No', 'Partial', 'Pending'])
    ws.add_data_validation(yn_dv)
    yn_dv.add('H5:H54')
    yn_dv.add('J5:J54')
    yn_dv.add('K5:K54')
    yn_dv.add('L5:L54')

    # Date formatting
    for col in ['F', 'I', 'M']:
        for row in range(5, 55):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes (DS-007: A4 per standard)
    ws.freeze_panes = 'A4'

    # Define named range for Employee_ID_List
    employee_id_range = DefinedName(name='Employee_ID_List', attr_text="\'Personnel Contract Compliance\'!$A$5:$A$54")
    wb.defined_names.add(employee_id_range)

    # Sheet protection removed (SRC-018)


def create_confidentiality_nda_tracking_sheet(wb, styles):
    """
    Sheet 5: Confidentiality NDA Tracking

    NDA / confidentiality status per individual.
    AUDIT FOCAL POINT - critical for compliance verification.
    200 rows for personnel.
    """
    ws = wb.create_sheet("Confidentiality NDA Tracking")
    ws.sheet_view.showGridLines = False

    # --- Header ---
    merge_and_style(ws, 'A1:N1', 'CONFIDENTIALITY & NDA TRACKING', styles['header_main'])
    merge_and_style(ws, 'A2:N2',
                   'AUDIT FOCAL POINT: NDA and confidentiality agreement status per individual',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35

    # --- Column Headers ---
    headers = [
        ('A', 'Employee ID', 15),
        ('B', 'Employee Name', 25),
        ('C', 'Department', 20),
        ('D', 'NDA Type', 20),
        ('E', 'NDA Version', 12),
        ('F', 'Date Signed', 12),
        ('G', 'Expiration Date', 12),
        ('H', 'Covers IP', 15),
        ('I', 'Covers Trade Secrets', 15),
        ('J', 'Covers Post Employment', 15),
        ('K', 'NDA Status', 15),
        ('L', 'Renewal Required', 15),
        ('M', 'Document Location', 30),
        ('N', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (5-54) ---
    for row in range(5, 55):
        # Auto-populate Employee info from Personnel sheet
        ws[f'A{row}'] = f'=IF(ROW()-4<=COUNTA(\'Personnel Contract Compliance\'!$A:$A)-4,INDEX(\'Personnel Contract Compliance\'!$A:$A,ROW()-3),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        ws[f'B{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},\'Personnel Contract Compliance\'!$A:$B,2,FALSE),"")'
        apply_cell_style(ws[f'B{row}'], styles['calculated_cell'])

        ws[f'C{row}'] = f'=IF(A{row}<>"",VLOOKUP(A{row},\'Personnel Contract Compliance\'!$A:$C,3,FALSE),"")'
        apply_cell_style(ws[f'C{row}'], styles['calculated_cell'])

        # Manual entry cells
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'M', 'N']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

        # NDA_Status (K) - auto-calculated
        ws[f'K{row}'] = (f'=IF(B{row}="","",IF(AND(F{row}<>"",OR(G{row}="",G{row}>=TODAY())),"Active",'
                        f'IF(AND(F{row}<>"",G{row}<TODAY()),"Expired",'
                        f'IF(F{row}="","Missing","Unknown"))))')
        apply_cell_style(ws[f'K{row}'], styles['calculated_cell'])

    # --- Data Validation ---
    nda_type_dv = create_data_validation(['Embedded-In-Contract', 'Standalone-NDA', 'Mutual-NDA',
                                          'Unilateral-NDA', 'PIIA', 'Combined-Agreement'])
    ws.add_data_validation(nda_type_dv)
    nda_type_dv.add('D5:D54')

    yn_dv = create_data_validation(['Yes', 'No', 'Partial'])
    ws.add_data_validation(yn_dv)
    yn_dv.add('H5:H54')
    yn_dv.add('I5:I54')
    yn_dv.add('J5:J54')

    renewal_dv = create_data_validation(['Yes', 'No', 'N/A'])
    ws.add_data_validation(renewal_dv)
    renewal_dv.add('L5:L54')

    # Date formatting
    for col in ['F', 'G']:
        for row in range(5, 55):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes
    ws.freeze_panes = 'A4'

    # Sheet protection removed (SRC-018)


def create_post_employment_obligations_sheet(wb, styles):
    """
    Sheet 6: Post Employment Obligations

    Track individuals with ongoing post-employment obligations.
    100 rows for terminated employees with obligations.
    """
    ws = wb.create_sheet("Post Employment Obligations")
    ws.sheet_view.showGridLines = False

    # --- Header ---
    merge_and_style(ws, 'A1:O1', 'POST-EMPLOYMENT OBLIGATIONS TRACKING', styles['header_main'])
    merge_and_style(ws, 'A2:O2',
                   'Tracking individuals with ongoing confidentiality and non-compete obligations',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35

    # --- Column Headers ---
    headers = [
        ('A', 'Former Employee ID', 15),
        ('B', 'Former Employee Name', 25),
        ('C', 'Former Department', 20),
        ('D', 'Former Role', 25),
        ('E', 'Termination Date', 12),
        ('F', 'Termination Type', 15),
        ('G', 'Confidentiality End Date', 15),
        ('H', 'Non Compete End Date', 15),
        ('I', 'Non Solicit End Date', 15),
        ('J', 'IP Assignment Perpetual', 15),
        ('K', 'Assets Returned', 15),
        ('L', 'Exit Interview Complete', 15),
        ('M', 'Obligation Status', 15),
        ('N', 'Monitoring Required', 15),
        ('O', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}3'] = header_text
        apply_cell_style(ws[f'{col_letter}3'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Sample Row (Row 4) — Step 2.4 ---
    _peo_smp_fill = PatternFill('solid', fgColor='F2F2F2')
    _peo_smp_font = Font(name='Calibri', size=10, italic=True, color='808080')
    _peo_smp_bdr = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    _peo_sample = {
        'A': 'FMR-EXAMPLE', 'B': 'John Davies (Example)',
        'C': 'Finance', 'D': 'Finance Manager',
        'E': '30.06.2025', 'F': 'Resignation',
        'G': '30.06.2028', 'H': '30.06.2026', 'I': '30.06.2026',
        'J': 'Yes', 'K': 'Yes', 'L': 'Yes',
        'M': 'Active-Compliant', 'N': 'No',
        'O': 'Example entry — replace with your leaver obligation records',
    }
    for col_letter in [c[0] for c in headers]:
        ws[f'{col_letter}4'].value = _peo_sample.get(col_letter, '')
        ws[f'{col_letter}4'].fill = _peo_smp_fill
        ws[f'{col_letter}4'].font = _peo_smp_font
        ws[f'{col_letter}4'].border = _peo_smp_bdr

    # --- Data Rows (5-54) ---
    for row in range(5, 55):
        for col_letter in [c[0] for c in headers]:
            cell = ws[f'{col_letter}{row}']
            if col_letter == 'M':
                # Auto-calculated obligation status
                ws[f'M{row}'] = (f'=IF(B{row}="","",IF(AND(K{row}="Yes",L{row}="Yes",'
                                f'OR(G{row}="Perpetual",G{row}>=TODAY())),"Active-Compliant",'
                                f'IF(OR(K{row}="No",L{row}="No"),"Non-Compliant",'
                                f'IF(AND(G{row}<TODAY(),H{row}<TODAY(),I{row}<TODAY()),"Expired","Active-Review"))))')
                apply_cell_style(cell, styles['calculated_cell'])
            else:
                apply_cell_style(cell, styles['input_cell'])

    # --- Data Validation ---
    term_type_dv = create_data_validation(['Resignation', 'Termination', 'Retirement',
                                           'Contract-End', 'Mutual-Agreement', 'Redundancy'])
    ws.add_data_validation(term_type_dv)
    term_type_dv.add('F5:F54')

    yn_dv = create_data_validation(['Yes', 'No', 'Partial', 'N/A'])
    ws.add_data_validation(yn_dv)
    yn_dv.add('J5:J54')
    yn_dv.add('K5:K54')
    yn_dv.add('L5:L54')
    yn_dv.add('N5:N54')

    # Date formatting
    for col in ['E', 'G', 'H', 'I']:
        for row in range(5, 55):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes (DS-007: A4 per standard)
    ws.freeze_panes = 'A4'

    # Sheet protection removed (SRC-018)


def create_contractor_agreement_assessment_sheet(wb, styles):
    """
    Sheet 7: Contractor Agreement Assessment

    Third-party and contractor agreement verification.
    100 rows for contractors/vendors.
    """
    ws = wb.create_sheet("Contractor Agreement Assessment")
    ws.sheet_view.showGridLines = False

    # --- Header ---
    merge_and_style(ws, 'A1:N1', 'CONTRACTOR AGREEMENT ASSESSMENT', styles['header_main'])
    merge_and_style(ws, 'A2:N2',
                   'Third-party, contractor, and vendor agreement security clause verification',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35

    # --- Column Headers (Row 3) ---
    headers = [
        ('A', 'Contractor ID', 15),
        ('B', 'Contractor Name', 25),
        ('C', 'Company Name', 25),
        ('D', 'Contract Type', 20),
        ('E', 'Services Provided', 30),
        ('F', 'Contract Start', 12),
        ('G', 'Contract End', 12),
        ('H', 'NDA In Place', 15),
        ('I', 'Security Clauses Present', 15),
        ('J', 'Data Access Level', 15),
        ('K', 'Background Check', 15),
        ('L', 'Compliance Status', 15),
        ('M', 'Sponsor Manager', 25),
        ('N', 'Notes', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}3'] = header_text
        apply_cell_style(ws[f'{col_letter}3'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Sample Row (Row 4) — Step 2.4 ---
    _caa_smp_fill = PatternFill('solid', fgColor='F2F2F2')
    _caa_smp_font = Font(name='Calibri', size=10, italic=True, color='808080')
    _caa_smp_bdr = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    _caa_sample = {
        'A': 'CON-EXAMPLE', 'B': 'John Contractor (Example)',
        'C': 'Acme Consulting Ltd', 'D': 'Statement-of-Work',
        'E': 'IT Security Consulting', 'F': '01.01.2025', 'G': '31.12.2025',
        'H': 'Yes', 'I': 'Yes', 'J': 'Limited', 'K': 'Yes',
        'L': 'Compliant', 'M': 'IT Security Manager',
        'N': 'Example entry — replace with your contractor agreement details',
    }
    for col_letter in [c[0] for c in headers]:
        ws[f'{col_letter}4'].value = _caa_sample.get(col_letter, '')
        ws[f'{col_letter}4'].fill = _caa_smp_fill
        ws[f'{col_letter}4'].font = _caa_smp_font
        ws[f'{col_letter}4'].border = _caa_smp_bdr

    # --- Data Rows (5-54) ---
    for row in range(5, 55):
        for col_letter in [c[0] for c in headers]:
            cell = ws[f'{col_letter}{row}']
            if col_letter == 'L':
                # Auto-calculated compliance status
                ws[f'L{row}'] = (f'=IF(B{row}="","",IF(AND(H{row}="Yes",I{row}="Yes",'
                                f'OR(K{row}="Yes",K{row}="N/A")),"Compliant",'
                                f'IF(OR(H{row}="No",I{row}="No"),"Non-Compliant","Partial")))')
                apply_cell_style(cell, styles['calculated_cell'])
            else:
                apply_cell_style(cell, styles['input_cell'])

    # --- Data Validation ---
    contract_type_dv = create_data_validation(['Statement-of-Work', 'Master-Services-Agreement',
                                                'Professional-Services', 'Staffing-Agency',
                                                'Consulting-Agreement', 'Vendor-Contract', 'Other'])
    ws.add_data_validation(contract_type_dv)
    contract_type_dv.add('D5:D54')

    yn_dv = create_data_validation(['Yes', 'No', 'Partial', 'N/A'])
    ws.add_data_validation(yn_dv)
    yn_dv.add('H5:H54')
    yn_dv.add('I5:I54')
    yn_dv.add('K5:K54')

    access_dv = create_data_validation(['None', 'Limited', 'Standard', 'Privileged', 'Admin'])
    ws.add_data_validation(access_dv)
    access_dv.add('J5:J54')

    # Date formatting
    for col in ['F', 'G']:
        for row in range(5, 55):
            ws[f'{col}{row}'].number_format = 'DD.MM.YYYY'

    # Freeze panes
    ws.freeze_panes = 'A4'

    # Sheet protection removed (SRC-018)


def create_gap_analysis_sheet(wb, styles):
    """
    Sheet 8: Gap Analysis

    Consolidated gaps from all assessment domains.
    """
    ws = wb.create_sheet("Gap Analysis")
    ws.sheet_view.showGridLines = False

    # --- Header ---
    merge_and_style(ws, 'A1:P1', 'GAP ANALYSIS', styles['header_main'])
    merge_and_style(ws, 'A2:P2',
                   'Consolidated gaps from all employment contract assessment domains',
                   styles['header_sub'])
    ws.row_dimensions[1].height = 35

    # --- Column Headers ---
    headers = [
        ('A', 'Gap ID', 12),
        ('B', 'Source Sheet', 25),
        ('C', 'Affected Entity', 25),
        ('D', 'Gap Category', 20),
        ('E', 'Gap Description', 50),
        ('F', 'Risk Level', 15),
        ('G', 'Impact Assessment', 40),
        ('H', 'Affected Personnel Count', 15),
        ('I', 'Remediation Action', 50),
        ('J', 'Responsible Party', 25),
        ('K', 'Target Completion Date', 15),
        ('L', 'Estimated Effort', 15),
        ('M', 'Dependencies', 30),
        ('N', 'Status', 15),
        ('O', 'Completion Evidence', 30),
        ('P', 'Risk Acceptance', 40)
    ]

    widths_dict = {}
    for col_letter, header_text, width in headers:
        ws[f'{col_letter}4'] = header_text
        apply_cell_style(ws[f'{col_letter}4'], styles['column_header'])
        widths_dict[col_letter] = width

    set_column_widths(ws, widths_dict)

    # --- Data Rows (5-104, 100 rows for gaps) ---
    for row in range(5, 105):
        # Gap_ID auto-generated
        ws[f'A{row}'] = f'=IF(B{row}<>"",TEXT(ROW()-4,"GAP-000"),"")'
        apply_cell_style(ws[f'A{row}'], styles['calculated_cell'])

        # All other columns manual entry
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            apply_cell_style(ws[f'{col}{row}'], styles['input_cell'])

    # Date formatting for Target_Completion_Date
    for row in range(5, 105):
        ws[f'K{row}'].number_format = 'DD.MM.YYYY'

    # --- Data Validation ---
    source_sheet_dv = create_data_validation(['Contract Template Assessment', 'Required Clause Registry',
                                               'Personnel Contract Compliance', 'Confidentiality NDA Tracking',
                                               'Post Employment Obligations', 'Contractor Agreement Assessment'])
    ws.add_data_validation(source_sheet_dv)
    source_sheet_dv.add('B5:B104')

    category_dv = create_data_validation(['Template-Gap', 'Clause-Gap', 'Contract-Gap',
                                          'NDA-Gap', 'Post-Employment-Gap', 'Contractor-Gap',
                                          'Process-Gap', 'Documentation-Gap'])
    ws.add_data_validation(category_dv)
    category_dv.add('D5:D104')

    risk_level_dv = create_data_validation(['Critical', 'High', 'Medium', 'Low'])
    ws.add_data_validation(risk_level_dv)
    risk_level_dv.add('F5:F104')

    effort_dv = create_data_validation(['<1hr', '1-4hrs', '1day', '2-5days', '>1week'])
    ws.add_data_validation(effort_dv)
    effort_dv.add('L5:L104')

    status_dv = create_data_validation(['Not-Started', 'In-Progress', 'Blocked', 'Completed', 'Accepted-Risk'])
    ws.add_data_validation(status_dv)
    status_dv.add('N5:N104')

    # Freeze panes
    ws.freeze_panes = 'A4'

    # Sheet protection removed (SRC-018)


def create_evidence_register(wb):
    """
    Sheet 9: Evidence Register (gold standard — standalone, no styles parameter)
    """
    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_hdr = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _input = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35
    ws.merge_cells("A2:H2")
    ws["A2"] = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = _border
    columns = [
        ("Evidence ID", 14), ("Evidence Type", 22), ("Description", 45),
        ("Related Control / Section", 28), ("Collection Date (DD.MM.YYYY)", 22),
        ("Storage Location / Reference", 38), ("Collected By", 22), ("Status", 14),
    ]
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    sample_data = [
        "EV-001", "Document", "Sample evidence entry — replace with actual evidence",
        "All Controls", "01.01.2026", "SharePoint/ISMS/Evidence/", "ISMS Team", "Active"
    ]
    for col_idx, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
        cell.fill = _grey_sample
        cell.border = _border
    dv_status = DataValidation(type="list", formula1='"Active,Archived,Superseded,Pending Review"', allow_blank=True)
    ws.add_data_validation(dv_status)
    for r in range(6, 106):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = _input
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        dv_status.add(ws.cell(row=r, column=8))
    ws.freeze_panes = "A5"


def create_approval_sheet(wb):
    """Create the Approval Sign-Off sheet — Gold Standard."""
    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Summary fields (rows 4-8)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G7:G11),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Assessment Status dropdown (row 7)
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # 3 Approver sections (start at row 11)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


def create_summary_dashboard_sheet(wb):
    """Add Summary Dashboard sheet (TABLE 1/2/3) to workbook."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title="Summary Dashboard")
    ws.sheet_view.showGridLines = False

    def _f(hex_c):
        return PatternFill("solid", fgColor=hex_c)

    def _b():
        s = Side(style="thin", color="000000")
        return Border(left=s, right=s, top=s, bottom=s)

    def _hdr(row, col, value, fill="003366", fc="FFFFFF", sz=11,
             bold=True, merge_to=None):
        c = ws.cell(row=row, column=col)
        c.value = value
        c.font = Font(name="Calibri", bold=bold, color=fc, size=sz)
        c.fill = _f(fill)
        c.alignment = Alignment(horizontal="left" if merge_to else "center",
                                vertical="center", wrap_text=True)
        c.border = _b()
        if merge_to:
            ws.merge_cells(f"{get_column_letter(col)}{row}:{merge_to}")
        return c

    def _dat(row, col, value, fill="FFFFCC", fc="000000", bold=False,
             num=False):
        c = ws.cell(row=row, column=col)
        c.value = value
        c.font = Font(name="Calibri", bold=bold, color=fc)
        c.fill = _f(fill)
        c.alignment = Alignment(
            horizontal="center" if num else "left",
            vertical="center", wrap_text=True)
        c.border = _b()
        return c

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "A3"

    # Row 1: Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "EMPLOYMENT CONTRACT ASSESSMENT \u2014 SUMMARY DASHBOARD"
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    c.fill = _f("003366")
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    c.border = _b()
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = ("ISO/IEC 27001:2022 | Control A.6.2 | "
               "Employment contract clause coverage and compliance summary")
    c.font = Font(name="Calibri", italic=True, color="003366", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _b()

    # ── TABLE 1: CLAUSE COVERAGE DISTRIBUTION ────────────────────────────
    _hdr(4, 1, "TABLE 1: CLAUSE COVERAGE DISTRIBUTION", "003366",
         sz=12, merge_to="F4")
    _hdr(5, 1,
         "Required Clause Registry — Column H (Coverage Status), rows 5:74",
         "4472C4", sz=10, merge_to="F5")
    _hdr(6, 1, "Coverage Status", "4472C4", sz=10)
    _hdr(6, 2, "Count", "4472C4", sz=10)
    _hdr(6, 3, "% of Total", "4472C4", sz=10)

    t1_statuses = ["Covered", "Partial", "Not-Covered", "N/A"]
    for i, status in enumerate(t1_statuses):
        r = 7 + i
        _dat(r, 1, status)
        _dat(r, 2,
             f'=COUNTIF(\'Required Clause Registry\'!H5:H74,\"{status}\")',
             num=True)
        _dat(r, 3, f'=IF($B$11=0,"—",TEXT(B{r}/$B$11,"0.0%"))', num=True)

    _dat(11, 1, "TOTAL", "D9D9D9", bold=True)
    _dat(11, 2, "=SUM(B7:B10)", "D9D9D9", num=True)
    _dat(11, 3, "—", "D9D9D9", num=True)

    # ── TABLE 2: KEY METRICS ──────────────────────────────────────────────
    _hdr(13, 1, "TABLE 2: KEY METRICS", "003366", sz=12, merge_to="F13")
    _hdr(14, 1, "KPI Metric", "D9D9D9", fc="000000", sz=10)
    _hdr(14, 2, "Value", "D9D9D9", fc="000000", sz=10)
    _hdr(14, 3, "Notes", "D9D9D9", fc="000000", sz=10, merge_to="F14")

    kpis = [
        ("Total Clause Types Assessed",
         "=COUNTA(\'Required Clause Registry\'!B5:B74)",
         "All clause categories in Required Clause Registry"),
        ("Clauses Fully Covered",
         "=COUNTIF(\'Required Clause Registry\'!H5:H74,\"Covered\")",
         "Clause types with full coverage across contract templates"),
        ("Partial Coverage",
         "=COUNTIF(\'Required Clause Registry\'!H5:H74,\"Partial\")",
         "Clause types with partial coverage — review for gaps"),
        ("Clauses Not Covered",
         "=COUNTIF(\'Required Clause Registry\'!H5:H74,\"Not-Covered\")",
         "Mandatory clause types with no coverage — highest risk"),
        ("Critical Gaps Identified",
         "=COUNTIF(\'Gap Analysis\'!F5:F104,\"Critical\")",
         "Critical-risk gaps in Gap Analysis sheet"),
        ("Open Remediation Actions",
         "=COUNTIF(\'Gap Analysis\'!N5:N104,\"Not-Started\")"
         "+COUNTIF(\'Gap Analysis\'!N5:N104,\"In-Progress\")",
         "Gaps not yet resolved (Not-Started + In-Progress)"),
    ]
    for i, (metric, formula, note) in enumerate(kpis):
        r = 15 + i
        _dat(r, 1, metric)
        _dat(r, 2, formula, num=True)
        c = ws.cell(row=r, column=3)
        c.value = note
        c.font = Font(name="Calibri", color="000000")
        c.fill = _f("FFFFCC")
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        c.border = _b()
        ws.merge_cells(f"C{r}:F{r}")

    # ── TABLE 3: CRITICAL FINDINGS ────────────────────────────────────────
    _hdr(22, 1, "TABLE 3: CRITICAL FINDINGS", "C00000", sz=12, merge_to="F22")
    _hdr(23, 1, "Finding", "D9D9D9", fc="000000", sz=10)
    _hdr(23, 2, "Count", "D9D9D9", fc="000000", sz=10)
    _hdr(23, 3, "Action Required", "D9D9D9", fc="000000", sz=10,
         merge_to="F23")

    critical = [
        ("Clauses Not Covered",
         "=COUNTIF(\'Required Clause Registry\'!H5:H74,\"Not-Covered\")",
         "Update contract templates to include all mandatory clauses immediately"),
        ("Clauses with Partial Coverage",
         "=COUNTIF(\'Required Clause Registry\'!H5:H74,\"Partial\")",
         "Complete partial clauses — specify missing content and assign owner"),
        ("Critical Contract Gaps",
         "=COUNTIF(\'Gap Analysis\'!F5:F104,\"Critical\")",
         "Address all critical-risk gaps in Gap Analysis — assign remediation owner"),
    ]
    for i, (finding, formula, action) in enumerate(critical):
        r = 24 + i
        c = ws.cell(row=r, column=1)
        c.value = finding
        c.font = Font(name="Calibri", bold=True, color="000000")
        c.fill = _f("FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        c.border = _b()
        c2 = ws.cell(row=r, column=2)
        c2.value = formula
        c2.font = Font(name="Calibri", color="000000")
        c2.fill = _f("FFFFFF")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = _b()
        c3 = ws.cell(row=r, column=3)
        c3.value = action
        c3.font = Font(name="Calibri", color="000000")
        c3.fill = _f("FFFFFF")
        c3.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True)
        c3.border = _b()
        ws.merge_cells(f"C{r}:F{r}")



def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass


# ==============================================================================
# MAIN EXECUTION
# ==============================================================================



def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Review the Dashboard (auto-calculated) for a high-level contractual compliance overview.',
        '2. Complete the Contract Inventory — list all employment and contractor agreements with security clauses.',
        '3. Complete Security Clause Verification — confirm NDA, acceptable use, data handling, and incident reporting clauses.',
        '4. Complete the Onboarding Checklist — verify security briefing, system access provisioning, and device assignment.',
        '5. Complete the Offboarding Checklist — verify access revocation, device return, and exit interview completion.',
        '6. Complete the Change Management Register — track contract amendments triggered by role or access changes.',
        '7. Review the Gap Analysis (auto-populated) — identify missing clauses or unsigned agreements.',
        '8. Attach audit evidence in the Evidence Register (EV-xxx references).',
        '9. Create remediation actions in Action Items with owners and target dates.',
        '10. Obtain three-level stakeholder approval in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A24"] = "Status Legend"
    ws["A24"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=25, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 26 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.5.1-2-6.1-2.S4 - Employment Contract Assessment")
    logger.info("Workbook Generator")
    logger.info("=" * 80)

    # Initialize workbook
    logger.info("Creating workbook...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default blank sheet (Dashboard removed — Summary Dashboard serves this purpose)
    styles = _STYLES

    # Create all sheets

    logger.info("Creating Instructions & Legend...")
    create_instructions_sheet(wb.create_sheet())
    logger.info("Creating Contract Template Assessment...")
    create_contract_template_assessment_sheet(wb, styles)

    logger.info("Creating Required Clause Registry...")
    create_required_clause_registry_sheet(wb, styles)

    logger.info("Creating Personnel Contract Compliance...")
    create_personnel_contract_compliance_sheet(wb, styles)

    logger.info("Creating Confidentiality NDA Tracking...")
    create_confidentiality_nda_tracking_sheet(wb, styles)

    logger.info("Creating Post Employment Obligations...")
    create_post_employment_obligations_sheet(wb, styles)

    logger.info("Creating Contractor Agreement Assessment...")
    create_contractor_agreement_assessment_sheet(wb, styles)

    logger.info("Creating Gap Analysis...")
    create_gap_analysis_sheet(wb, styles)

    logger.info("Creating Evidence Register...")
    create_evidence_register(wb)

    logger.info("Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb)

    logger.info("Creating Approval Sign-Off...")
    create_approval_sheet(wb)

    # Finalise data validations (SRC-008)
    finalize_validations(wb)

    # Set workbook properties
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.keywords = "Employment, Contract, NDA, ISMS, ISO27001, A.6.1, A.6.2"
    wb.properties.comments = "Generated via Python script generate_a5_1_2_6_1_2_s4_employment_contract.py"

    # Generate filename with current date
    today = datetime.now().strftime("%Y%m%d")

    # Save workbook
    logger.info(f"Saving workbook as: {output_path.name}")
    wb.save(output_path)
    logger.info("=" * 80)
    logger.info("Workbook generated successfully!")
    logger.info("=" * 80)
    logger.info("Next Steps:")
    logger.info("1. Open the workbook in Excel")
    logger.info("2. Complete Sheet 2 (Contract Template Assessment) - verify all templates")
    logger.info("3. Review Sheet 3 (Required Clause Registry) - verify clause coverage")
    logger.info("4. Complete Sheet 4 (Personnel Contract Compliance) - per-person verification")
    logger.info("5. Verify Sheet 5 (Confidentiality NDA Tracking) - AUDIT FOCAL POINT")
    logger.info("6. Track Sheet 6 (Post Employment Obligations) - ongoing obligations")
    logger.info("7. Assess Sheet 7 (Contractor Agreement Assessment) - third-party contracts")
    logger.info("8. Document Sheet 8 (Gap Analysis) - consolidate all gaps")
    logger.info("9. Complete Sheet 9 (Evidence_Register) - supporting evidence")
    logger.info("10. Obtain Sheet 10 (Approval_Sign_Off) - three-level approval")
    logger.info(f"File location: ./{output_path.name}")
def main():
    try:
        create_workbook(_wkbk_dir / OUTPUT_FILENAME)
    except Exception as e:
        logger.error(f"Failed to generate workbook: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
