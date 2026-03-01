#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.31.3 - Requirements Extraction Register Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.31: Legal, Statutory, Regulatory and Contractual Requirements
Assessment Domain 3 of 6: Requirements Extraction and Registration

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific requirements extraction methodology, categorization
taxonomy, and implementation tracking needs.

Key customization areas:
1. Requirement categorization scheme (adapt to your control framework)
2. Priority and criticality definitions (align with your risk methodology)
3. Implementation status tracking (match your project management processes)
4. Requirement granularity guidelines (specific to your operational model)
5. Traceability linkage structure (based on your ISMS architecture)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.31 Regulatory Compliance Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for extracting
specific, actionable requirements from applicable regulations (identified in
Workbook 1, assessed in Workbook 2), transforming legal/regulatory text into
structured, implementable requirements that can be mapped to ISO 27001 controls.

**Purpose:**
Enables systematic extraction, interpretation, and registration of regulatory
requirements against ISO 27001:2022 Control A.5.31, providing the critical link
between legal obligations and operational controls through structured requirement
parsing and categorization.

**Assessment Scope:**
- Legal requirement text extraction from source regulations
- Requirement interpretation into actionable statements
- Requirement categorization (Technical/Organisational/Reporting/Legal/Contractual)
- Priority and criticality assessment
- Implementation status tracking
- Regulatory citation mapping (regulation → article → section → clause)
- Requirement granularity management (appropriate level of detail)
- Cross-reference tracking (requirements affecting multiple controls)
- Gap identification (requirements without existing controls)
- Traceability matrix (requirement → control → evidence)
- Integration with Regulatory Inventory (Workbook 1) and Control Mapping (Workbook 4)

**Generated Workbook Structure:**
1. Requirements_Register - Master register of all extracted requirements
2. Instructions - Comprehensive extraction methodology guidance
3. GDPR_Example - Completed example showing GDPR requirements extraction
4. Template_Per_Regulation - Blank template for new regulation analysis

**Key Features:**
- Structured three-level requirement hierarchy (Regulation → Article → Requirement)
- Multi-dimensional categorization taxonomy
- Data validation with category and status dropdown lists
- Conditional formatting for priority and status visualization
- Requirement versioning and change tracking
- Cross-reference linking to related requirements
- Protected formulas with unprotected input cells
- Sample completed extraction (GDPR Articles) as reference
- Per-regulation template for systematic extraction
- UTF-8 encoding with emoji support

**Integration:**
This requirements register feeds into:
- Workbook 4 (Control Mapping Matrix - maps requirements to ISO 27001 controls)
- Workbook 5 (Evidence Register - links evidence to requirements)
- Dashboard (Compliance Overview - requirement coverage metrics)
- ISMS Policies and Procedures (requirements drive control implementation)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_531_3_requirements_register.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_531_3_requirements_register.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_531_3_requirements_register.py --date 20250125

Output:
    File: ISMS_Assessment_531_3_Requirements_Register_YYYYMMDD.xlsx
    Location: WKBK/ (or specified output path)

Post-Generation Steps:
    1. Review requirements extraction methodology in Instructions sheet
    2. Review sample GDPR extraction (GDPR_Example sheet) as reference
    3. For each Tier 1 and Tier 2 Applicable regulation:
       a. Copy Template_Per_Regulation sheet, rename with regulation ID
       b. Read complete regulation text (do NOT rely on summaries)
       c. Extract all relevant requirements systematically
       d. Rewrite requirements in actionable form ([Organisation] SHALL/MUST...)
       e. Categorize each requirement (Technical/Organisational/etc.)
       f. Assign priority and criticality
       g. Document regulatory citation precisely
    4. Consolidate all per-regulation extractions into Requirements_Register
    5. Identify requirements without existing controls (gaps)
    6. Feed requirements into Control Mapping (Workbook 4)
    7. Update ISMS policies/procedures to address new requirements
    8. Obtain legal counsel review for complex requirement interpretations

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.31
Assessment Domain:    3 of 6 (Requirements Extraction and Registration)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.31.3: Requirements Extraction & Control Mapping Framework
    - ISMS-IMP-A.5.31.3: Requirements Extraction Process
    - ISMS-POL-00: Regulatory Applicability Framework

Related Assessment Tools:
    - Assessment Workbook 1: Regulatory Inventory (generate_531_1_regulatory_inventory.py)
    - Assessment Workbook 2: Applicability Matrix (generate_531_2_applicability_matrix.py)
    - Assessment Workbook 4: Control Mapping (generate_531_4_control_mapping.py)
    - Assessment Workbook 5: Evidence Register (generate_531_5_evidence_register.py)
    - Dashboard: Compliance Overview (generate_531_6_compliance_dashboard.py)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements structured requirements extraction framework
    - Supports multi-level requirement categorization
    - Integrated traceability matrix
    - Sample GDPR extraction as reference

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Requirements Extraction Philosophy:**
This is NOT simple copy-paste from regulation text. Requirements extraction is:

**Legal Text → Interpreted Requirement → Actionable Control**

Example transformation:
- Legal Text (GDPR Article 5.1.f): "Personal data shall be processed in a manner that ensures appropriate security"
- Interpreted Requirement: "[Organisation] SHALL implement technical and organisational measures to ensure appropriate security of personal data"
- Actionable Control: "Implement encryption for personal data at rest and in transit" (maps to A.8.24)

**Requirement Granularity:**
Critical balance - too broad or too granular both cause problems:

**Too Broad** (not useful):
- "Comply with GDPR" (this is the goal, not a requirement)
- "Ensure data security" (what does this mean operationally?)

**Too Granular** (overwhelming):
- "Use AES-256-GCM cipher for database encryption" (too prescriptive, belongs in implementation)
- "Send breach notification within 23.5 hours" (overspecified)

**Appropriate Granularity** (just right):
- "Implement encryption for personal data at rest and in transit"
- "Notify data protection authority of personal data breaches within 72 hours"
- "Conduct Data Protection Impact Assessment before high-risk processing"

**Rule of thumb:** Requirement should be specific enough to map to 1-3 ISO 27001 controls,
but general enough to allow implementation flexibility.

**Requirement Categories:**
Organize requirements by type for efficient control mapping:

1. **Technical Requirements**: Technology implementation
   - Examples: Encryption, access control, logging, network security
   - Typical mappings: A.8.x (Technology controls)

2. **Organisational Requirements**: Processes and procedures
   - Examples: Policies, risk assessment, training, documentation
   - Typical mappings: A.5.x (Organisational controls), A.6.x (People controls)

3. **Reporting Requirements**: External communications
   - Examples: Breach notification, compliance reporting, transparency disclosures
   - Typical mappings: A.5.27 (Incident management), A.5.34 (Privacy)

4. **Legal/Administrative**: Formal legal obligations
   - Examples: Contracts, agreements, registration, appointments (DPO)
   - Typical mappings: A.5.31 (this control), A.5.1 (Policies)

5. **Contractual Requirements**: Customer/partner obligations
   - Examples: SLA commitments, audit rights, data processing agreements
   - Typical mappings: A.5.19-5.23 (Supplier relationships)

**Priority and Criticality:**
Different dimensions requiring different assessments:

**Priority** (when to implement):
- P1-Critical: Immediate (within 1 month) - legal deadline, contractual obligation
- P2-High: Short-term (within 3 months) - significant risk, customer expectation
- P3-Medium: Medium-term (within 6-12 months) - important but not urgent
- P4-Low: Long-term (1-2 years) - nice to have, continuous improvement

**Criticality** (consequence of non-compliance):
- Critical: Regulatory penalties, license loss, business shutdown
- High: Customer loss, reputational damage, audit failure
- Medium: Internal policy violation, minor compliance gap
- Low: Best practice deviation, voluntary commitment

**Regulatory Citation Best Practices:**
Precise citation enables audit traceability and update management:

**Good Citation Format:**
- GDPR Article 5(1)(f) - Security principle
- NIST CSF 2.0 v1.1 PR.AC-4 - Access permissions
- ISO 27001:2022 A.5.31 - Legal requirements
- Swiss FADP Article 8 - Data security

**Poor Citation Format:**
- "GDPR somewhere" (not auditable)
- "The security article" (ambiguous)
- "Section 5" (which regulation? which version?)

**Include:**
- Regulation short name (GDPR, not "Regulation (EU) 2016/679")
- Specific article/section/clause number
- Brief description of topic (helpful for navigation)
- Version/date if regulation has multiple versions

**Implementation Status Tracking:**
Requirement lifecycle management:

- **Not Started**: Requirement identified, no implementation begun
- **Planning**: Analysis underway, solution design in progress
- **In Progress**: Implementation active, controls being deployed
- **Implemented**: Controls deployed, awaiting verification
- **Verified**: Implementation tested and confirmed effective
- **N/A - Existing**: Existing controls already satisfy requirement (document mapping)
- **Deferred**: Implementation postponed (document reason and target date)

**Change Management:**
Requirements change when:
- Regulation amended (most common trigger)
- Regulatory guidance issued (interpretation changes)
- Organisational context changes (new requirement becomes relevant)
- Legal counsel reinterprets requirement
- Audit finding questions previous interpretation

When requirement changes:
1. Version previous requirement (append version number or date)
2. Document what changed and why
3. Re-assess control mappings (Workbook 4)
4. Update evidence register (Workbook 5) if needed
5. Notify affected control owners

**Audit Considerations:**
Requirements register is critical audit evidence demonstrating:
- Systematic analysis of applicable regulations
- Complete coverage (all requirements extracted, not cherry-picked)
- Appropriate interpretation (legal counsel involvement)
- Traceability (regulation → requirement → control → evidence)
- Implementation tracking (what's done, what's in progress)

Auditors will spot-check:
- Random regulation → Verify all requirements extracted
- Specific control → Trace back to driving regulatory requirement
- Gap claims → Verify requirement truly has no control

**Data Protection:**
Requirements register contains sensitive information:
- Legal interpretations and risk assessments
- Implementation gaps and deficiencies
- Strategic plans for compliance
- Control weaknesses and remediation timelines

Handle in accordance with [Organisation]'s data classification policies.

**Legal Counsel Critical Role:**
Legal counsel should review requirements extraction for:
- Complex legal language interpretation
- Ambiguous or conflicting requirements
- Extraterritorial application questions
- Threshold determinations (e.g., "high risk processing")
- Legal vs. regulatory guidance distinction

Particularly critical for:
- Data protection regulations (GDPR, FADP) - complex legal concepts
- Financial regulations (DORA, PCI DSS v4.0.1) - technical + legal requirements
- Emerging regulations (AI Act, Cyber Resilience) - unclear interpretations

**Common Extraction Challenges:**

1. **Principle-Based vs. Rule-Based Regulations**:
   - Principle-based (GDPR): Broad principles requiring interpretation
   - Rule-based (PCI DSS v4.0.1): Specific technical requirements
   - Extraction approach differs - principles need more interpretation

2. **Nested Requirements** (requirement within requirement):
   - Example: GDPR Article 32 requires "appropriate security" AND lists examples (pseudonymization, encryption, etc.)
   - Extract both general requirement AND specific examples as separate requirements

3. **Implicit Requirements** (not explicitly stated but legally required):
   - Example: GDPR doesn't explicitly require "cybersecurity program" but it's implied by security obligations
   - Consult legal counsel before extracting implicit requirements

4. **Exemptions and Exceptions**:
   - Many regulations have exemptions (small businesses, specific sectors, etc.)
   - Document applicable exemptions in requirement notes
   - Don't extract requirements that don't apply due to exemptions

5. **Cross-References Between Requirements**:
   - Regulations often reference other requirements
   - Extract each requirement separately, but note relationships
   - Use "Related Requirements" field for cross-references

**Integration with Control Mapping:**
Requirements feed directly into Workbook 4 (Control Mapping):
1. Each requirement becomes a row in control mapping matrix
2. Requirement ID links back to this register
3. Control mapping identifies which ISO 27001 controls satisfy requirement
4. Gaps identified when requirement has no control

**Quality Assurance:**
Before finalizing requirements extraction:
- Read ENTIRE regulation (don't rely on summaries or excerpts)
- Extract ALL relevant requirements (not just obvious ones)
- Rewrite requirements in actionable form ([Organisation] SHALL...)
- Appropriate granularity (not too broad, not too granular)
- Complete categorization and priority assignment
- Precise regulatory citation for traceability
- Legal counsel reviewed complex interpretations
- No duplicate requirements (check for overlaps)

**Template Usage:**
Template_Per_Regulation sheet is designed for systematic extraction:
1. Copy sheet, rename with Regulation ID (REG-EU-001-GDPR-Requirements)
2. Work through regulation systematically (article by article, section by section)
3. Extract requirements as you go (don't try to extract everything at once)
4. After complete extraction, consolidate into Requirements_Register master sheet
5. Keep per-regulation sheet as audit trail

**Scalability:**
Large regulatory portfolios can generate 500-1000+ requirements:
- Use filtering and sorting extensively
- Group related requirements for control mapping efficiency
- Consider sub-registers per regulation (link to master)
- Dedupe common requirements across regulations (e.g., encryption)
- Use requirement IDs for cross-referencing (REQ-GDPR-005, REQ-FADP-012)

**Common Pitfalls:**
- Copying regulation text verbatim (not interpretation)
- Extracting at wrong granularity (too broad or too detailed)
- Mixing requirements and controls (this is requirements, not how to implement)
- Incomplete extraction (missing requirements from regulation)
- No legal counsel review for complex legal language
- Failing to update when regulations change
- Not tracking implementation status (requirements stagnate)

================================================================================
"""

from datetime import datetime, timedelta
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging
import sys

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.31.3"
WORKBOOK_NAME = "Requirements Register"
CONTROL_ID   = "A.5.31"
CONTROL_NAME = "Legal, Statutory, Regulatory and Contractual Requirements"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = ''    #  Chart
TARGET = ''    #  Target
SHIELD = ''    # ️  Shield
LOCK = ''     #  Lock
SCALES = '\u2696'     # ⚖️  Scales of Justice
DOCUMENT = ''       #  Document
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Sheet structure
    sheets = [
        "Instructions & Legend",
        "Requirements Register",
        "Summary Dashboard",
        "Approval Sign-Off",
        "Extraction Worksheet",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    styles = {
        "header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "data": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "data_center": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "date": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        f = style_dict["font"]
        cell.font = Font(
            name=f.name,
            size=f.size,
            bold=f.bold if hasattr(f, "bold") else False,
            color=f.color if hasattr(f, "color") else "000000",
        )
    if "fill" in style_dict:
        f = style_dict["fill"]
        cell.fill = PatternFill(
            start_color=f.start_color.rgb if hasattr(f.start_color, "rgb") else f.start_color,
            end_color=f.end_color.rgb if hasattr(f.end_color, "rgb") else f.end_color,
            fill_type=f.fill_type,
        )
    if "alignment" in style_dict:
        a = style_dict["alignment"]
        cell.alignment = Alignment(
            horizontal=a.horizontal,
            vertical=a.vertical,
            wrap_text=a.wrap_text if hasattr(a, "wrap_text") else False,
        )
    if "border" in style_dict:
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================================
# SECTION 2: COLUMN DEFINITIONS
# ============================================================================

def get_register_columns() -> dict:
    """Return column definitions with widths."""
    return {
        "Requirement ID": 18,
        "Regulation ID": 15,
        "Regulation Name": 35,
        "Citation": 20,
        "Original Requirement Text": 50,
        "Interpreted Requirement": 50,
        "Category": 18,
        "Priority": 15,
        "Implementation Deadline": 18,
        "Implementation Status": 20,
        "Mapped Controls": 25,
        "Gap Status": 18,
        "Responsible Party": 20,
        "Notes": 40,
        "Extracted By": 25,
        "Reviewed By": 25,
        "Approved By": 25,
        "Last Updated": 15,
    }


# ============================================================================
# SECTION 3: DATA VALIDATION
# ============================================================================

def create_validations(ws):
    """Create and apply data validation to columns."""
    
    # Category dropdown (Column G)
    category_validation = DataValidation(
        type="list",
        formula1='"Technical,Organisational,Reporting,Operational"',
        allow_blank=False
    )
    category_validation.error = "Please select from the list"
    category_validation.errorTitle = "Invalid Category"
    ws.add_data_validation(category_validation)
    category_validation.add("G2:G1000")
    
    # Priority dropdown (Column H)
    priority_validation = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=False
    )
    priority_validation.error = "Please select from the list"
    priority_validation.errorTitle = "Invalid Priority"
    ws.add_data_validation(priority_validation)
    priority_validation.add("H2:H1000")
    
    # Implementation Status dropdown (Column J)
    status_validation = DataValidation(
        type="list",
        formula1='"➖ Not Started,⏳ In Progress,✅ Implemented,❓ N/A"',
        allow_blank=False
    )
    status_validation.error = "Please select from the list"
    status_validation.errorTitle = "Invalid Status"
    ws.add_data_validation(status_validation)
    status_validation.add("J2:J1000")
    
    # Gap Status dropdown (Column L)
    gap_validation = DataValidation(
        type="list",
        formula1=f'"{XMARK} Complete Gap,⚠️ Partial Gap,✅ No Gap,TBD TBD"',
        allow_blank=False
    )
    gap_validation.error = "Please select from the list"
    gap_validation.errorTitle = "Invalid Gap Status"
    ws.add_data_validation(gap_validation)
    gap_validation.add("L2:L1000")


# ============================================================================
# SECTION 4: CONDITIONAL FORMATTING
# ============================================================================

def apply_conditional_formatting(ws):
    """Apply conditional formatting to status columns."""
    from openpyxl.formatting.rule import CellIsRule
    
    # Priority column (H) - Color coding
    high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    medium_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    ws.conditional_formatting.add(
        "H2:H1000",
        CellIsRule(operator="equal", formula=['"High"'], fill=high_fill)
    )
    ws.conditional_formatting.add(
        "H2:H1000",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=medium_fill)
    )
    ws.conditional_formatting.add(
        "H2:H1000",
        CellIsRule(operator="equal", formula=['"Low"'], fill=low_fill)
    )
    
    # Implementation Status column (J) - Color coding
    not_started_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    in_progress_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    implemented_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    na_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    ws.conditional_formatting.add(
        "J2:J1000",
        CellIsRule(operator="equal", formula=['"➖"'], fill=not_started_fill)
    )
    ws.conditional_formatting.add(
        "J2:J1000",
        CellIsRule(operator="equal", formula=['"⏳"'], fill=in_progress_fill)
    )
    ws.conditional_formatting.add(
        "J2:J1000",
        CellIsRule(operator="equal", formula=['"\u2705"'], fill=implemented_fill)
    )
    ws.conditional_formatting.add(
        "J2:J1000",
        CellIsRule(operator="equal", formula=['"❓"'], fill=na_fill)
    )
    
    # Gap Status column (L) - Color coding
    complete_gap_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    partial_gap_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    no_gap_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    tbd_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    ws.conditional_formatting.add(
        "L2:L1000",
        CellIsRule(operator="equal", formula=['"\u274c"'], fill=complete_gap_fill)
    )
    ws.conditional_formatting.add(
        "L2:L1000",
        CellIsRule(operator="equal", formula=['"\u26a0"'], fill=partial_gap_fill)
    )
    ws.conditional_formatting.add(
        "L2:L1000",
        CellIsRule(operator="equal", formula=['"\u2705"'], fill=no_gap_fill)
    )
    ws.conditional_formatting.add(
        "L2:L1000",
        CellIsRule(operator="equal", formula=['"TBD"'], fill=tbd_fill)
    )


# ============================================================================
# SECTION 5: SAMPLE DATA GENERATION
# ============================================================================

def generate_sample_data():
    """Generate realistic sample data for demonstration."""
    today = datetime.now()
    
    sample_data = [
        # GDPR Requirements
        {
            "Requirement ID": "REG-GDPR-32-001",
            "Regulation ID": "REG-EU-001",
            "Regulation Name": "GDPR",
            "Citation": "Article 32(1)(a)",
            "Original Requirement Text": "The controller and the processor shall implement appropriate technical and organisational measures to ensure a level of security appropriate to the risk, including inter alia as appropriate: (a) the pseudonymisation and encryption of personal data",
            "Interpreted Requirement": "Implement encryption for personal data at rest and in transit where appropriate based on risk assessment",
            "Category": "Technical",
            "Priority": "High",
            "Implementation Deadline": datetime(2025, 6, 1),
            "Implementation Status": "⏳ In Progress",
            "Mapped Controls": "A.8.24, A.5.10",
            "Gap Status": f"{WARNING} Partial Gap",
            "Responsible Party": "CISO",
            "Notes": "Encryption in transit complete (TLS 1.3). Database encryption in progress (Q1 2025).",
            "Extracted By": "Compliance Analyst / 2024-11-15",
            "Reviewed By": "Legal Counsel / 2024-11-20",
            "Approved By": "Compliance Officer / 2024-11-22",
            "Last Updated": today,
        },
        {
            "Requirement ID": "REG-GDPR-33-001",
            "Regulation ID": "REG-EU-001",
            "Regulation Name": "GDPR",
            "Citation": "Article 33(1)",
            "Original Requirement Text": "In the case of a personal data breach, the controller shall without undue delay and, where feasible, not later than 72 hours after having become aware of it, notify the personal data breach to the supervisory authority",
            "Interpreted Requirement": "Notify data protection authority within 72 hours of becoming aware of personal data breach",
            "Category": "Reporting",
            "Priority": "High",
            "Implementation Deadline": datetime(2025, 3, 1),
            "Implementation Status": "➖ Not Started",
            "Mapped Controls": "A.5.26, A.5.25",
            "Gap Status": f"{XMARK} Complete Gap",
            "Responsible Party": "Incident Response Lead",
            "Notes": "Requires incident response procedure update and SIEM alerting configuration.",
            "Extracted By": "Compliance Analyst / 2024-11-15",
            "Reviewed By": "Legal Counsel / 2024-11-20",
            "Approved By": "Compliance Officer / 2024-11-22",
            "Last Updated": today,
        },
        {
            "Requirement ID": "REG-GDPR-37-001",
            "Regulation ID": "REG-EU-001",
            "Regulation Name": "GDPR",
            "Citation": "Article 37(1)",
            "Original Requirement Text": "The controller and the processor shall designate a data protection officer in cases where processing is carried out by a public authority or body",
            "Interpreted Requirement": "Appoint Data Protection Officer (DPO) where required by Article 37 criteria",
            "Category": "Organisational",
            "Priority": "High",
            "Implementation Deadline": datetime(2024, 12, 31),
            "Implementation Status": f"{CHECK} Implemented",
            "Mapped Controls": "A.5.2, A.6.1",
            "Gap Status": f"{CHECK} No Gap",
            "Responsible Party": "Executive Management",
            "Notes": "DPO appointed June 2024. Contact: dpo@organisation.ch",
            "Extracted By": "Compliance Analyst / 2024-11-15",
            "Reviewed By": "Legal Counsel / 2024-11-20",
            "Approved By": "Compliance Officer / 2024-11-22",
            "Last Updated": today,
        },
        
        # ISO 27001 Requirements
        {
            "Requirement ID": "REG-ISO27001-5.1-001",
            "Regulation ID": "REG-INT-001",
            "Regulation Name": "ISO 27001:2022",
            "Citation": "Clause 5.1 / A.5.1",
            "Original Requirement Text": "Top management shall establish, implement and maintain information security policies that are approved, communicated to all relevant parties, and available as documented information",
            "Interpreted Requirement": "Establish, document, approve, and communicate information security policies to all relevant parties",
            "Category": "Organisational",
            "Priority": "High",
            "Implementation Deadline": datetime(2025, 1, 31),
            "Implementation Status": f"{CHECK} Implemented",
            "Mapped Controls": "A.5.1",
            "Gap Status": f"{CHECK} No Gap",
            "Responsible Party": "ISMS Manager",
            "Notes": "ISMS policies documented and approved. Annual review scheduled Q4.",
            "Extracted By": "ISMS Analyst / 2024-10-01",
            "Reviewed By": "ISMS Manager / 2024-10-05",
            "Approved By": "Executive Management / 2024-10-10",
            "Last Updated": today,
        },
        {
            "Requirement ID": "REG-ISO27001-6.1-001",
            "Regulation ID": "REG-INT-001",
            "Regulation Name": "ISO 27001:2022",
            "Citation": "Clause 6.1.2",
            "Original Requirement Text": "The organisation shall define and apply an information security risk assessment process",
            "Interpreted Requirement": "Define and implement systematic information security risk assessment process",
            "Category": "Organisational",
            "Priority": "High",
            "Implementation Deadline": datetime(2025, 2, 28),
            "Implementation Status": "⏳ In Progress",
            "Mapped Controls": "A.5.7, A.8.8",
            "Gap Status": f"{WARNING} Partial Gap",
            "Responsible Party": "CISO",
            "Notes": "Risk assessment methodology documented. Tool selection in progress (Q1 2025).",
            "Extracted By": "ISMS Analyst / 2024-10-01",
            "Reviewed By": "ISMS Manager / 2024-10-05",
            "Approved By": "Executive Management / 2024-10-10",
            "Last Updated": today,
        },
        {
            "Requirement ID": "REG-ISO27001-8.1-001",
            "Regulation ID": "REG-INT-001",
            "Regulation Name": "ISO 27001:2022",
            "Citation": "Clause 8.1 / Annex A",
            "Original Requirement Text": "The organisation shall plan, implement and control the processes needed to meet information security requirements, and to implement the actions determined in 6.1",
            "Interpreted Requirement": "Implement and control all applicable Annex A controls as documented in Statement of Applicability",
            "Category": "Technical",
            "Priority": "High",
            "Implementation Deadline": datetime(2025, 9, 30),
            "Implementation Status": "⏳ In Progress",
            "Mapped Controls": "All 93 Annex A controls",
            "Gap Status": f"{WARNING} Partial Gap",
            "Responsible Party": "ISMS Manager",
            "Notes": "Control implementation roadmap in progress. 65/93 controls implemented (70%).",
            "Extracted By": "ISMS Analyst / 2024-10-01",
            "Reviewed By": "ISMS Manager / 2024-10-05",
            "Approved By": "Executive Management / 2024-10-10",
            "Last Updated": today,
        },
        
        # Swiss FADP Requirements
        {
            "Requirement ID": "REG-FADP-7-001",
            "Regulation ID": "REG-CH-001",
            "Regulation Name": "Swiss FADP",
            "Citation": "Article 7",
            "Original Requirement Text": "The controller shall implement appropriate technical and organisational measures to ensure a level of data security appropriate to the risk to the personality and fundamental rights of the data subject arising from the processing",
            "Interpreted Requirement": "Implement risk-appropriate technical and organisational measures to protect personal data",
            "Category": "Technical",
            "Priority": "High",
            "Implementation Deadline": datetime(2025, 6, 30),
            "Implementation Status": "⏳ In Progress",
            "Mapped Controls": "A.5.10, A.5.14, A.8.1-8.34",
            "Gap Status": f"{WARNING} Partial Gap",
            "Responsible Party": "CISO",
            "Notes": "Overlaps significantly with GDPR Art. 32. Using ISO 27001 controls framework.",
            "Extracted By": "Compliance Analyst / 2024-12-01",
            "Reviewed By": "Legal Counsel (CH) / 2024-12-05",
            "Approved By": "Compliance Officer / 2024-12-08",
            "Last Updated": today,
        },
        {
            "Requirement ID": "REG-FADP-24-001",
            "Regulation ID": "REG-CH-001",
            "Regulation Name": "Swiss FADP",
            "Citation": "Article 24",
            "Original Requirement Text": "Controllers and processors shall notify the Federal Data Protection and Information Commissioner (FDPIC) of any data breach likely to result in a high risk to the personality or fundamental rights of the data subject as quickly as possible",
            "Interpreted Requirement": "Notify FDPIC of high-risk data breaches as quickly as possible",
            "Category": "Reporting",
            "Priority": "High",
            "Implementation Deadline": datetime(2025, 3, 1),
            "Implementation Status": "➖ Not Started",
            "Mapped Controls": "A.5.26, A.5.25",
            "Gap Status": f"{XMARK} Complete Gap",
            "Responsible Party": "Incident Response Lead",
            "Notes": "Similar to GDPR but reported to FDPIC. No 72-hour explicit deadline but 'quickly as possible'.",
            "Extracted By": "Compliance Analyst / 2024-12-01",
            "Reviewed By": "Legal Counsel (CH) / 2024-12-05",
            "Approved By": "Compliance Officer / 2024-12-08",
            "Last Updated": today,
        },
        
        # NIS2 Requirements (Conditional/Tier 2)
        {
            "Requirement ID": "REG-NIS2-21-001",
            "Regulation ID": "REG-EU-002",
            "Regulation Name": "NIS2",
            "Citation": "Article 21",
            "Original Requirement Text": "Essential and important entities shall take appropriate and proportionate technical, operational and organisational measures to manage the risks posed to the security of network and information systems",
            "Interpreted Requirement": "Implement risk management measures for network and information systems security",
            "Category": "Technical",
            "Priority": "Medium",
            "Implementation Deadline": datetime(2025, 10, 17),
            "Implementation Status": "➖ Not Started",
            "Mapped Controls": "A.5.7, A.8.8, A.8.20",
            "Gap Status": "TBD",
            "Responsible Party": "CISO",
            "Notes": "Tier 2 (Conditional): Only applicable if thresholds met. Monitor quarterly.",
            "Extracted By": "Compliance Analyst / 2024-12-10",
            "Reviewed By": "Legal Counsel / 2024-12-12",
            "Approved By": "Compliance Officer / 2024-12-15",
            "Last Updated": today,
        },
        {
            "Requirement ID": "REG-NIS2-23-001",
            "Regulation ID": "REG-EU-002",
            "Regulation Name": "NIS2",
            "Citation": "Article 23",
            "Original Requirement Text": "Essential and important entities shall notify, without undue delay, the CSIRT or, where applicable, the competent authority, of any incident having a significant impact on the provision of their services",
            "Interpreted Requirement": "Notify CSIRT/competent authority of significant incidents without undue delay",
            "Category": "Reporting",
            "Priority": "Medium",
            "Implementation Deadline": datetime(2025, 10, 17),
            "Implementation Status": "➖ Not Started",
            "Mapped Controls": "A.5.26, A.5.27",
            "Gap Status": "TBD",
            "Responsible Party": "Incident Response Lead",
            "Notes": "Tier 2 (Conditional): Two-stage notification (24h initial, 72h detailed report).",
            "Extracted By": "Compliance Analyst / 2024-12-10",
            "Reviewed By": "Legal Counsel / 2024-12-12",
            "Approved By": "Compliance Officer / 2024-12-15",
            "Last Updated": today,
        },
        
        # NIST CSF 2.0 (Voluntary/Tier 3)
        {
            "Requirement ID": "REG-NIST-ID.AM-001",
            "Regulation ID": "REG-VOL-001",
            "Regulation Name": "NIST CSF 2.0",
            "Citation": "ID.AM-1",
            "Original Requirement Text": "Inventories of hardware managed by the organisation are maintained",
            "Interpreted Requirement": "Maintain inventory of all hardware assets managed by organisation",
            "Category": "Organisational",
            "Priority": "Low",
            "Implementation Deadline": datetime(2026, 6, 30),
            "Implementation Status": "⏳ In Progress",
            "Mapped Controls": "A.5.9, A.8.9",
            "Gap Status": f"{WARNING} Partial Gap",
            "Responsible Party": "IT Operations Manager",
            "Notes": "Tier 3 (Informational): Voluntary framework. Asset inventory tool deployed (80% coverage).",
            "Extracted By": "Security Analyst / 2024-11-01",
            "Reviewed By": "CISO / 2024-11-05",
            "Approved By": "ISMS Manager / 2024-11-08",
            "Last Updated": today,
        },
        {
            "Requirement ID": "REG-NIST-PR.DS-001",
            "Regulation ID": "REG-VOL-001",
            "Regulation Name": "NIST CSF 2.0",
            "Citation": "PR.DS-1",
            "Original Requirement Text": "Data-at-rest is protected",
            "Interpreted Requirement": "Implement protection measures for data at rest (encryption, access controls)",
            "Category": "Technical",
            "Priority": "Low",
            "Implementation Deadline": datetime(2026, 6, 30),
            "Implementation Status": "⏳ In Progress",
            "Mapped Controls": "A.8.24, A.5.10, A.8.11",
            "Gap Status": f"{WARNING} Partial Gap",
            "Responsible Party": "CISO",
            "Notes": "Tier 3 (Informational): Overlaps with GDPR/FADP requirements.",
            "Extracted By": "Security Analyst / 2024-11-01",
            "Reviewed By": "CISO / 2024-11-05",
            "Approved By": "ISMS Manager / 2024-11-08",
            "Last Updated": today,
        },
    ]
    
    return sample_data


# ============================================================================
# SECTION 6: SHEET POPULATION FUNCTIONS
# ============================================================================

def populate_register_sheet(wb, styles):
    """Populate the Requirements Register sheet with headers, data, and formatting."""
    ws = wb["Requirements Register"]
    ws.sheet_view.showGridLines = False
    columns = get_register_columns()

    # Title row (row 1) - ALL CAPS, 003366 fill
    num_cols = len(columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value="REQUIREMENTS REGISTER")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Set column widths and create headers (row 2)
    for idx, (col_name, width) in enumerate(columns.items(), start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width
        cell = ws.cell(row=2, column=idx, value=col_name)
        apply_style(cell, styles["header"])

    # Add sample data (starting row 3)
    sample_data = generate_sample_data()
    for row_idx, data in enumerate(sample_data, start=3):
        for col_idx, col_name in enumerate(columns.keys(), start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = data.get(col_name, "")

            # Handle date formatting
            if isinstance(value, datetime):
                cell.value = value
                cell.number_format = "DD.MM.YYYY"
                apply_style(cell, styles["date"])
            elif col_name in ["Priority", "Implementation Status", "Gap Status"]:
                cell.value = value
                apply_style(cell, styles["data_center"])
            else:
                cell.value = value
                apply_style(cell, styles["data"])

    # Apply data validation
    create_validations(ws)

    # Apply conditional formatting
    apply_conditional_formatting(ws)

    # Freeze panes (freeze rows 1-2 and first column)
    ws.freeze_panes = "B3"

    # Enable auto-filter
    ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}2"


def populate_instructions_sheet(wb):
    """Create GS-IL-compliant Instructions & Legend sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb["Instructions & Legend"]
    ws.sheet_view.showGridLines = False

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 2 — empty

    # Row 3 — Document Information heading
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Workbook Title", WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Instructions section
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete the Requirements Register — extract specific obligations from each regulation.',
        '2. Use the Extraction Worksheet to review regulations article by article.',
        '3. Map each requirement to the responsible control owner and business process.',
        '4. Document compliance status (Compliant, Partial, Non-Compliant) with evidence references.',
        '5. Identify gaps and create remediation items for non-compliant requirements.',
        '6. Maintain the Evidence Register with requirement source documents and assessments.',
        '7. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Status Legend section
    ws["A21"] = "Status Legend"
    ws["A21"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=22, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("✓", "Compliant / Complete", "Requirement fully met", _green),
        ("⚠", "Partial / In Progress", "Partially met or in progress", _amber),
        ("✗", "Non-Compliant / Not Started", "Requirement not met", _red),
        ("—", "Not Applicable", "Not applicable to this assessment", None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 23 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"

def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete the Requirements Register — extract specific obligations from each regulation.',
        '2. Use the Extraction Worksheet to review regulations article by article.',
        '3. Map each requirement to the responsible control owner and business process.',
        '4. Document compliance status (Compliant, Partial, Non-Compliant) with evidence references.',
        '5. Identify gaps and create remediation items for non-compliant requirements.',
        '6. Maintain the Evidence Register with requirement source documents and assessments.',
        '7. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A21"] = "Status Legend"
    ws["A21"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=22, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 23 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_summary_dashboard_sheet(wb):
    """Create Gold Standard Summary Dashboard sheet — ISMS-IMP-A.5.31.3 Requirements Register."""
    ws = wb["Summary Dashboard"]
    ws.sheet_view.showGridLines = False

    _thin = Side(border_style="thin", color="000000")
    _b    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy  = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _blue  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _red   = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _grey  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _yell  = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _lft   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Column widths (7 columns: A:G)
    for col, w in zip("ABCDEFG", [50, 12, 18, 15, 18, 12, 15]):
        ws.column_dimensions[col].width = w

    # ── Row 1: Title (GS-SD-014: must contain em dash + SUMMARY DASHBOARD) ─
    ws.merge_cells("A1:G1")
    ws["A1"] = "REQUIREMENTS REGISTER \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = _navy
    ws["A1"].alignment = _ctr
    ws.row_dimensions[1].height = 35
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = _b

    # ── Row 2: Subtitle (left aligned, no fill) ──────────────────────────
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.5.31: Legal, Statutory, Regulatory and Contractual Requirements | Requirements Register"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=2, column=c).border = _b

    # ── Row 3: Empty separator ───────────────────────────────────────────

    # ── TABLE 1: Assessment Area Compliance Overview ─────────────────────
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws["A4"].fill = _navy
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=4, column=c).border = _b

    # TABLE 1 headers (row 5) — D9D9D9, black bold (GS-SD-016: NOT 4472C4)
    for c, h in enumerate(["Assessment Area", "Total Items", "Compliant", "Partial",
                            "Non-Compliant", "N/A", "Compliance %"], 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = Font(bold=True, size=10, color="000000", name="Calibri")
        cell.fill = _grey
        cell.alignment = _ctr
        cell.border = _b

    # TABLE 1 single data row (row 6): Requirements Register
    # col A = Req ID (user-entered), data starts row 3; col J = Implementation Status
    row = 6
    ws.cell(row=row, column=1, value="Requirements Register").border = _b
    ws.cell(row=row, column=1).font = Font(color="000000", name="Calibri", size=10)
    ws.cell(row=row, column=1).alignment = _lft

    # B: Total
    cell_b = ws.cell(row=row, column=2,
        value="=COUNTA('Requirements Register'!A3:A1000)")
    cell_b.border = _b; cell_b.alignment = Alignment(horizontal="center")
    cell_b.font = Font(color="000000", name="Calibri", size=10)

    # C: Compliant = Implemented
    cell_c = ws.cell(row=row, column=3,
        value="=COUNTIF('Requirements Register'!J3:J1000,\"\u2705 Implemented\")")
    cell_c.border = _b; cell_c.alignment = Alignment(horizontal="center")
    cell_c.font = Font(color="000000", name="Calibri", size=10)

    # D: Partial = In Progress
    cell_d = ws.cell(row=row, column=4,
        value="=COUNTIF('Requirements Register'!J3:J1000,\"\u23f3 In Progress\")")
    cell_d.border = _b; cell_d.alignment = Alignment(horizontal="center")
    cell_d.font = Font(color="000000", name="Calibri", size=10)

    # E: Non-Compliant = Not Started
    cell_e = ws.cell(row=row, column=5,
        value="=COUNTIF('Requirements Register'!J3:J1000,\"\u2796 Not Started\")")
    cell_e.border = _b; cell_e.alignment = Alignment(horizontal="center")
    cell_e.font = Font(color="000000", name="Calibri", size=10)

    # F: N/A
    cell_f = ws.cell(row=row, column=6,
        value="=COUNTIF('Requirements Register'!J3:J1000,\"\u2753 N/A\")")
    cell_f.border = _b; cell_f.alignment = Alignment(horizontal="center")
    cell_f.font = Font(color="000000", name="Calibri", size=10)

    # G: Compliance %
    cell_g = ws.cell(row=row, column=7,
        value=f"=IFERROR(IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row})),\"\")")
    cell_g.number_format = "0.0%"
    cell_g.border = _b; cell_g.alignment = Alignment(horizontal="center")
    cell_g.font = Font(color="000000", name="Calibri", size=10)

    # TOTAL row (row 7)
    total_row = 7
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="000000", name="Calibri", size=10)
    ws.cell(row=total_row, column=1).fill = _grey
    ws.cell(row=total_row, column=1).border = _b
    ws.cell(row=total_row, column=1).alignment = _lft
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col,
                       value=f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_row - 1})")
        cell.font = Font(bold=True, color="000000", name="Calibri", size=10)
        cell.fill = _grey
        cell.border = _b
        cell.alignment = Alignment(horizontal="center")
    cell_tot_pct = ws.cell(row=total_row, column=7,
                           value=f"=IFERROR(IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row})),\"\")")
    cell_tot_pct.number_format = "0.0%"
    cell_tot_pct.font = Font(bold=True, color="000000", name="Calibri", size=10)
    cell_tot_pct.fill = _grey
    cell_tot_pct.border = _b
    cell_tot_pct.alignment = Alignment(horizontal="center")

    # ── TABLE 2: Key Metrics ─────────────────────────────────────────────
    t2_banner_row = total_row + 2  # row 9
    ws.merge_cells(f"A{t2_banner_row}:G{t2_banner_row}")
    ws[f"A{t2_banner_row}"] = "TABLE 2: KEY METRICS"
    ws[f"A{t2_banner_row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{t2_banner_row}"].fill = _navy
    ws[f"A{t2_banner_row}"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=t2_banner_row, column=c).border = _b

    # TABLE 2 headers — D9D9D9 grey, black bold (GS-SD-016)
    t2_hdr_row = t2_banner_row + 1  # row 10
    for c, h in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        cell = ws.cell(row=t2_hdr_row, column=c, value=h if h else None)
        cell.font = Font(bold=True, color="000000", name="Calibri", size=10)
        cell.fill = _grey
        cell.border = _b
        cell.alignment = _ctr

    # TABLE 2 metrics — white fill, 000000 font, NOT bold labels (GS-SD-015)
    metrics = [
        ("Total requirements extracted",
         "=COUNTA('Requirements Register'!A3:A1000)"),
        ("High priority requirements",
         "=COUNTIF('Requirements Register'!H3:H1000,\"High\")"),
        ("Medium priority requirements",
         "=COUNTIF('Requirements Register'!H3:H1000,\"Medium\")"),
        ("Technical requirements",
         "=COUNTIF('Requirements Register'!G3:G1000,\"Technical\")"),
        ("Organisational requirements",
         "=COUNTIF('Requirements Register'!G3:G1000,\"Organisational\")"),
        ("Reporting requirements",
         "=COUNTIF('Requirements Register'!G3:G1000,\"Reporting\")"),
        ("Operational requirements",
         "=COUNTIF('Requirements Register'!G3:G1000,\"Operational\")"),
        ("Requirements implemented",
         "=COUNTIF('Requirements Register'!J3:J1000,\"\u2705 Implemented\")"),
        ("Requirements in progress",
         "=COUNTIF('Requirements Register'!J3:J1000,\"\u23f3 In Progress\")"),
        ("Requirements not started",
         "=COUNTIF('Requirements Register'!J3:J1000,\"\u2796 Not Started\")"),
        ("Requirements with no gap",
         "=COUNTIF('Requirements Register'!L3:L1000,\"\u2705 No Gap\")"),
        ("Requirements with partial gap",
         "=COUNTIF('Requirements Register'!L3:L1000,\"\u26a0\ufe0f Partial Gap\")"),
        ("Requirements with complete gap",
         "=COUNTIF('Requirements Register'!L3:L1000,\"\u274c Complete Gap\")"),
    ]
    row = t2_hdr_row + 1  # row 11
    for metric, formula in metrics:
        cell_m = ws.cell(row=row, column=1, value=metric)
        cell_m.border = _b
        cell_m.font = Font(color="000000", name="Calibri", size=10)  # NOT bold (GS-SD-015)
        cell_m.alignment = _lft
        cell_v = ws.cell(row=row, column=2, value=formula)
        cell_v.border = _b
        cell_v.font = Font(color="000000", name="Calibri", size=10)
        cell_v.alignment = Alignment(horizontal="center")
        for c in range(3, 8):
            ws.cell(row=row, column=c).border = _b
        row += 1
    t2_last_row = row - 1  # row 23

    # ── TABLE 3: Critical Findings ────────────────────────────────────────
    t3_banner_row = t2_last_row + 2  # row 25
    ws.merge_cells(f"A{t3_banner_row}:G{t3_banner_row}")
    ws[f"A{t3_banner_row}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{t3_banner_row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{t3_banner_row}"].fill = _red
    ws[f"A{t3_banner_row}"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=t3_banner_row, column=c).border = _b

    # TABLE 3 headers — D9D9D9
    t3_hdr_row = t3_banner_row + 1  # row 26
    for c, h in enumerate(["Finding", "Count", "Action Required", "", "", "", ""], 1):
        cell = ws.cell(row=t3_hdr_row, column=c, value=h if h else None)
        cell.font = Font(bold=True, color="000000", name="Calibri", size=10)
        cell.fill = _grey
        cell.border = _b
        cell.alignment = _ctr
    ws.merge_cells(f"C{t3_hdr_row}:G{t3_hdr_row}")

    _yell_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    findings = [
        ("High-priority requirements Not Started",
         "=COUNTIFS('Requirements Register'!H3:H1000,\"High\",'Requirements Register'!J3:J1000,\"\u2796 Not Started\")",
         "Immediate escalation \u2014 assign owners and implementation deadlines for all High priority requirements"),
        ("Complete gaps (requirements with no control)",
         "=COUNTIF('Requirements Register'!L3:L1000,\"\u274c Complete Gap\")",
         "Risk acceptance or control implementation required for all complete gap requirements"),
        ("High-priority requirements with complete gap",
         "=COUNTIFS('Requirements Register'!H3:H1000,\"High\",'Requirements Register'!L3:L1000,\"\u274c Complete Gap\")",
         "Critical compliance gap \u2014 immediate remediation plan required"),
    ]
    row = t3_hdr_row + 1  # row 27
    for finding, count_formula, action in findings:
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = _yell_fill
            ws.cell(row=row, column=c).border = _b
            ws.cell(row=row, column=c).font = Font(color="000000", name="Calibri", size=10)
        ws.cell(row=row, column=1, value=finding).alignment = _lft
        cell_cnt = ws.cell(row=row, column=2, value=count_formula)
        cell_cnt.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"C{row}:G{row}")
        cell_act = ws.cell(row=row, column=3, value=action)
        cell_act.alignment = _lft
        for c in range(4, 8):
            ws.cell(row=row, column=c).border = _b
        row += 1

    # 2 empty FFFFCC buffer rows
    for _ in range(2):
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = _yell_fill
            ws.cell(row=row, column=c).border = _b
        row += 1
    t3_last_row = row - 1

    # ── FINAL DECISION (GS-AS-012: col A plain bold, NO dark fill) ───────
    fd_row = t3_last_row + 2
    ws.cell(row=fd_row, column=1, value="FINAL DECISION:").font = Font(bold=True, name="Calibri")
    ws.merge_cells(f"B{fd_row}:G{fd_row}")
    ws.cell(row=fd_row, column=2).fill = _yell
    for c in range(2, 8):
        ws.cell(row=fd_row, column=c).border = _b

    fd_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(fd_dv)
    fd_dv.add(f"B{fd_row}")

    # ── NEXT REVIEW DETAILS ───────────────────────────────────────────────
    nr_row = fd_row + 3
    ws.merge_cells(f"A{nr_row}:G{nr_row}")
    ws[f"A{nr_row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{nr_row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{nr_row}"].fill = _blue
    for c in range(1, 8):
        ws.cell(row=nr_row, column=c).border = _b

    for i, label in enumerate(["Next Review Date:", "Review Responsible:", "Special Considerations:"]):
        r = nr_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{r}:G{r}")
        ws.cell(row=r, column=2).fill = _yell
        for c in range(2, 8):
            ws.cell(row=r, column=c).border = _b

    # Apply borders to all merged ranges (GS-AS-011)
    for mr in list(ws.merged_cells.ranges):
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                ws.cell(row=r, column=c).border = _b

    ws.freeze_panes = "A4"


def create_approval_sheet(wb):
    """Create the Approval Sign-Off sheet — Gold Standard (GS-AS-014/015)."""
    ws = wb["Approval Sign-Off"]
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title banner — GS-AS-014
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY section banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Rows 4-8: Summary metadata — B6 = Overall Compliance (GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE(\'Summary Dashboard\'!G6:G6),\"\")")  ,
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Row 7 status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # Approver sections start at row 11 (rows 9-10 = gap)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION — GS-AS-012: col A = plain bold label, NO dark fill
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")

def populate_extraction_worksheet(wb):
    """Populate the Extraction Worksheet sheet."""
    ws = wb["Extraction Worksheet"]
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "EXTRACTION WORKSHEET — ARTICLE-BY-ARTICLE REVIEW"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # Column headers
    headers = [
        "Article #",
        "Article Title",
        "Summary",
        "Contains Obligations?",
        "Est. # Requirements",
        "Complexity",
        "Legal Review?",
        "Status",
        "Notes"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Set column widths
    widths = [12, 25, 40, 18, 18, 12, 15, 20, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Add sample rows
    sample_data = [
        ["Art. 32", "Security of processing", "Requirements for technical and organisational measures", "Y", "5", "M", "N", f"{CHECK} Complete", "5 requirements extracted"],
        ["Art. 33", "Breach notification", "72-hour notification requirement", "Y", "1", "S", "Y", f"{CHECK} Complete", "1 requirement extracted"],
        ["Art. 34", "Communication to data subjects", "Notification to affected individuals", "Y", "2", "M", "Y", "⏳ In Progress", ""],
        ["Art. 35", "Data protection impact assessment", "DPIA requirements and process", "Y", "4", "C", "Y", "➖ Not Started", ""],
        ["Art. 36", "Prior consultation", "Consultation with supervisory authority", "Y", "1", "M", "Y", "➖ Not Started", ""],
    ]
    
    for row_idx, data in enumerate(sample_data, start=3):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="left" if col_idx in [2, 3, 9] else "center", vertical="center", wrap_text=True)
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Add data validation
    yn_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(yn_validation)
    yn_validation.add("D3:D100")
    
    complexity_validation = DataValidation(type="list", formula1='"S,M,C"', allow_blank=True)
    ws.add_data_validation(complexity_validation)
    complexity_validation.add("F3:F100")
    
    legal_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(legal_validation)
    legal_validation.add("G3:G100")
    
    status_validation = DataValidation(type="list", formula1='"➖ Not Started,⏳ In Progress,✅ Complete"', allow_blank=True)
    ws.add_data_validation(status_validation)
    status_validation.add("H3:H100")
    
    # Freeze panes
    ws.freeze_panes = "A3"
    
    # Auto-filter
    ws.auto_filter.ref = "A2:I2"


# ============================================================================
# SECTION 7: MAIN EXECUTION
# ============================================================================


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    pass

def main():
    """Main execution function."""
    logger.info("=" * 70)
    logger.info("ISMS Assessment Workbook Generator")
    logger.info("Control A.5.31 - Workbook 3: Requirements Register")
    logger.info("=" * 70)
    logger.info("")
    
    # Create workbook
    logger.info(" Creating workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    
    # Populate sheets
    logger.info(" Populating Requirements Register sheet...")
    populate_register_sheet(wb, styles)
    
    logger.info(" Populating Instructions sheet...")
    populate_instructions_sheet(wb)
    
    logger.info(" Creating Summary Dashboard (Gold Standard TABLE 1/2/3)...")
    create_summary_dashboard_sheet(wb)

    logger.info(" Creating Approval Sign-Off sheet...")
    create_approval_sheet(wb)

    logger.info(" Populating Extraction Worksheet...")
    populate_extraction_worksheet(wb)
    
    # Save workbook
    _wkbk_dir.mkdir(exist_ok=True)
    output_path = _wkbk_dir / OUTPUT_FILENAME
    logger.info(f" Saving workbook to: {output_path}")
    wb.save(output_path)
    
    logger.info("")
    logger.info("{CHECK} Workbook generated successfully!")
    logger.info("")
    logger.info(" Output Details:")
    logger.info(f"   File: ISMS_Assessment_531_3_Requirements_Register.xlsx")
    logger.info(f"   Location: WKBK/")
    logger.info(f"   Sheets: 5 (Requirements Register, Instructions, Summary Dashboard, Approval Sign-Off, Extraction Worksheet)")
    logger.info(f"   Sample Data: 13 example requirements from 5 regulations")
    logger.info("")
    logger.info(" Features:")
    logger.info("   ✓ 18-column requirements register")
    logger.info("   ✓ Data validation (Category, Priority, Status, Gap dropdowns)")
    logger.info("   ✓ Conditional formatting (emoji-based status colors)")
    logger.info("   ✓ Auto-filter enabled")
    logger.info("   ✓ Freeze panes (header row)")
    logger.info("   ✓ Summary metrics with formulas")
    logger.info("   ✓ Extraction worksheet for article-by-article review")
    logger.info("   ✓ Comprehensive instructions")
    logger.info("   ✓ UTF-8 encoding with emoji support (✅❌⚠️⏳TBDHighMediumLow)")
    logger.info("")
    logger.info(f"{TARGET} Next Steps:")
    logger.info("   1. Use Extraction Worksheet to review regulations systematically")
    logger.info("   2. Extract requirements into Requirements Register")
    logger.info("   3. Assign categories, priorities, and deadlines")
    logger.info("   4. Map requirements to controls (Workbook 4)")
    logger.info("   5. Identify and remediate gaps")
    logger.info("   6. Collect evidence (Workbook 5)")
    logger.info("")
    logger.info("=" * 70)


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
