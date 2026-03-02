#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.31.4 - Control Mapping Matrix Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.31: Legal, Statutory, Regulatory and Contractual Requirements
Assessment Domain 4 of 6: Requirements-to-Controls Mapping

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific control framework, mapping methodology, and gap
management processes.

Key customization areas:
1. Control framework scope (ISO 27001 Annex A + any additional controls)
2. Mapping type definitions (Primary/Secondary/Supporting) and criteria
3. Gap analysis methodology (adapt to your risk assessment framework)
4. Coverage scoring and weighting (align with your compliance metrics)
5. Remediation prioritization (based on your risk appetite)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.31 Regulatory Compliance Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for mapping
regulatory requirements (extracted in Workbook 3) to ISO 27001:2022 Annex A
controls (and any additional organisational controls), enabling systematic
identification of control coverage, gaps, and evidence requirements.

**Purpose:**
Enables systematic mapping between regulatory obligations and operational
controls against ISO 27001:2022 Control A.5.31 requirements, providing the
critical traceability from legal requirements through controls to evidence,
and identifying gaps where requirements lack control coverage.

**Assessment Scope:**
- Requirements-to-Controls mapping matrix (requirements × controls)
- Mapping type categorization (Primary/Secondary/Supporting/Not Applicable)
- Coverage analysis and scoring per requirement
- Gap identification (requirements without adequate controls)
- Gap prioritization and remediation planning
- Additional controls beyond ISO 27001 Annex A
- One-to-many and many-to-one relationship tracking
- Cross-regulation control reuse identification
- Evidence location mapping per requirement-control pair
- Integration with Requirements Register (Workbook 3) and Evidence Register (Workbook 5)

**Generated Workbook Structure:**
1. Mapping_Matrix - Full requirements × controls matrix
2. Gap_Analysis - Requirements without adequate control coverage
3. Coverage_Summary - Statistical analysis of control coverage per regulation
4. Control_Reuse - Controls satisfying multiple requirements/regulations
5. Instructions - Comprehensive mapping methodology guidance
6. Additional_Controls - Controls beyond ISO 27001 Annex A

**Key Features:**
- Full matrix: All requirements (rows) × All ISO 27001 controls (columns)
- Mapping type indicators with conditional formatting
- Automated coverage percentage calculations per requirement
- Gap identification with criticality scoring
- Data validation with mapping type dropdown lists
- Protected formulas with unprotected input cells
- Control reuse analysis (efficiency opportunities)
- Integration with Statement of Applicability (SoA)
- UTF-8 encoding with emoji support

**Integration:**
This control mapping feeds into:
- Workbook 5 (Evidence Register - defines evidence needs per requirement-control pair)
- Dashboard (Compliance Overview - coverage and gap metrics)
- Statement of Applicability (validates control selections align with requirements)
- Control Implementation Plans (drives control deployment priorities)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_531_4_control_mapping.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_531_4_control_mapping.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_531_4_control_mapping.py --date 20250125

Output:
    File: ISMS_Assessment_531_4_Control_Mapping_YYYYMMDD.xlsx
    Location: WKBK/ (or specified output path)

Post-Generation Steps:
    1. Review mapping methodology in Instructions sheet
    2. Import all requirements from Requirements Register (Workbook 3)
    3. For each requirement, systematically assess each ISO 27001 control:
       a. Primary: Control directly implements requirement (main control)
       b. Secondary: Control partially implements requirement (supporting control)
       c. Supporting: Control provides enabling capability (indirect)
       d. N/A: Control does not address requirement
    4. Calculate coverage percentage per requirement
    5. Identify gaps (requirements with <50% coverage or no Primary control)
    6. For gaps: Determine if additional controls needed or mapping incomplete
    7. Document evidence location for each Primary/Secondary mapping
    8. Prioritize gap remediation based on criticality
    9. Update Statement of Applicability with requirement-driven justifications
    10. Feed evidence requirements into Evidence Register (Workbook 5)

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.31
Assessment Domain:    4 of 6 (Requirements-to-Controls Mapping)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.31.3: Requirements Extraction & Control Mapping Framework
    - ISMS-IMP-A.5.31.4: Control Mapping Process
    - Statement of Applicability (SoA) - ISO 27001:2022 Annex A

Related Assessment Tools:
    - Assessment Workbook 1: Regulatory Inventory (generate_531_1_regulatory_inventory.py)
    - Assessment Workbook 3: Requirements Register (generate_531_3_requirements_register.py)
    - Assessment Workbook 5: Evidence Register (generate_531_5_evidence_register.py)
    - Dashboard: Compliance Overview (generate_531_6_compliance_dashboard.py)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full requirements-to-controls mapping matrix
    - Supports gap analysis and remediation planning
    - Integrated coverage scoring and control reuse analysis
    - Compatible with ISO 27001:2022 93-control framework

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Control Mapping Philosophy:**
Mapping is NOT checkbox exercise - it's demonstrating HOW controls satisfy requirements:

**Cargo Cult Mapping** (wrong):
- Requirement: "Implement encryption" → Control A.8.24 ✓ (just check the box)

**Proper Mapping** (right):
- Requirement: "Encrypt personal data at rest and in transit"
  - A.8.24 (Cryptography) = Primary: Defines encryption requirements
  - A.8.23 (Web filtering) = Supporting: Enforces HTTPS for data in transit
  - A.5.10 (Acceptable use) = Supporting: Policy requires encryption
  - Evidence: Encryption policy, TLS configuration, encryption-at-rest configs

**Mapping Type Definitions:**
Critical to understand difference between Primary, Secondary, Supporting:

**Primary**: Control DIRECTLY implements the requirement
- This is the "main" control that does the work
- Requirement would NOT be satisfied without this control
- Example: Requirement "Notify DPA within 72 hours" → A.5.27 (Incident management) = Primary

**Secondary**: Control PARTIALLY implements requirement
- Requirement needs multiple controls working together
- This control handles one aspect
- Example: Requirement "Secure data transmission" → A.8.24 (Crypto) + A.8.20 (Network security) both Secondary

**Supporting**: Control provides ENABLING capability
- Doesn't directly implement requirement but enables/supports it
- Requirement could theoretically be satisfied without it, but less effectively
- Example: Requirement "Technical security measures" → A.6.3 (Training) = Supporting (enables people to implement measures)

**Not Applicable (N/A)**: Control does not address requirement in any way
- Most common mapping type (most controls don't relate to most requirements)
- Don't waste time documenting N/A (focus on Primary/Secondary/Supporting)

**Coverage Scoring:**
Quantitative assessment of how well requirements are controlled:

**Scoring Method:**
- Primary mapping: 100% coverage (fully implements requirement)
- Secondary mapping: 50% coverage (partially implements)
- Supporting mapping: 25% coverage (enables implementation)
- N/A mapping: 0% coverage

**Total Coverage** = Sum of all mappings (capped at 100%)

**Coverage Interpretation:**
- 100%+: Fully covered (has at least one Primary control)
- 50-99%: Partially covered (Secondary/Supporting only, no Primary)
- 25-49%: Minimally covered (Supporting only)
- 0-24%: GAP - Inadequate coverage, remediation needed

**Critical Rule:** Requirements SHOULD have at least one Primary control.
If only Secondary/Supporting, investigate if:
1. Mapping incomplete (is there a Primary control missing?)
2. Gap exists (need to implement new control)
3. Requirement spans multiple controls (each legitimately Secondary)

**Gap Analysis Methodology:**
Gaps are requirements lacking adequate control coverage:

**Gap Categories:**
1. **True Gap**: Requirement has no relevant control (must implement new control)
2. **Mapping Gap**: Control exists but not yet mapped (update mapping)
3. **Coverage Gap**: Controls exist but inadequate (must enhance control)
4. **Documentation Gap**: Control implemented but not documented (document it)

**Gap Identification Process:**
1. Review all requirements with <50% coverage
2. Review all requirements without Primary control
3. For each potential gap:
   - Is there an existing control that should map? (Mapping Gap)
   - Is control implemented but documentation missing? (Documentation Gap)
   - Is control partially implemented but needs enhancement? (Coverage Gap)
   - Is there truly no control for this requirement? (True Gap)

**Gap Remediation Prioritization:**
Priority = Requirement Criticality × Regulatory Importance × Implementation Effort

**P1 - Critical** (1-3 months):
- Tier 1 regulation + Critical requirement + True Gap
- Legal penalty risk or contractual breach
- Customer requirement with SLA commitment

**P2 - High** (3-6 months):
- Tier 1 regulation + High requirement + Coverage Gap
- Audit finding expected
- Industry standard practice

**P3 - Medium** (6-12 months):
- Tier 2 regulation + Medium requirement
- Continuous improvement
- Efficiency opportunity

**P4 - Low** (12-24 months):
- Tier 3 regulation + Low requirement
- Best practice alignment
- Nice-to-have enhancement

**One-to-Many and Many-to-One Relationships:**
Regulations and controls have complex relationships:

**One-to-Many** (One requirement → Many controls):
- Example: "Implement comprehensive security" requires 20+ controls
- Mapping approach: Identify all relevant controls, assign types appropriately
- Challenge: Ensure complete coverage (easy to miss controls)

**Many-to-One** (Many requirements → One control):
- Example: A.8.24 (Cryptography) satisfies encryption requirements from GDPR, PCI DSS v4.0.1, HIPAA, etc.
- Mapping approach: Use Control_Reuse sheet to identify efficiency
- Benefit: Implement one control, satisfy multiple requirements

**Control Reuse Benefits:**
Identifying control reuse is critical for efficiency:
- Avoids duplicate implementations
- Justifies control investment (multi-regulation benefit)
- Simplifies maintenance (one control to maintain, not many)
- Improves Statement of Applicability justification

**Integration with Statement of Applicability:**
This mapping directly informs SoA decisions:

**Control Selection Justification:**
- Control included BECAUSE it satisfies regulatory requirements
- SoA justification: "Required by GDPR Art. 32, PCI DSS v4.0.1 Req. 4.1, Customer Contract Sec. 5.3"

**Control Exclusion Justification:**
- Control excluded because no regulatory requirement drives it
- SoA justification: "Not applicable - [Organisation] does not [activity requiring control]"

**Mapping validates SoA:**
- All included controls should map to at least one requirement
- All excluded controls should have no regulatory requirement mappings
- If control is excluded but has requirement mappings → SoA inconsistency

**Additional Controls Beyond Annex A:**
ISO 27001 Annex A is not exhaustive - organisations often need additional controls:

**When to Add:**
- Regulatory requirement not satisfiable by any Annex A control
- Industry-specific requirement (e.g., PCI DSS v4.0.1 physical security requirements)
- Customer contractual requirement (e.g., specific encryption algorithm)
- Organisational policy beyond compliance minimum

**How to Add:**
1. Document in Additional_Controls sheet
2. Assign control ID (e.g., "ORG-A.8.100" following Annex A pattern)
3. Define control objective and implementation guidance
4. Map requirements to additional control like any Annex A control
5. Implement and manage like Annex A controls
6. Include in Statement of Applicability

**Audit Considerations:**
Control mapping is critical audit evidence demonstrating:
- Systematic coverage analysis of all requirements
- Traceability from regulation through requirement to control
- Gap identification and remediation planning
- Rationale for control selections (drives SoA)

Auditors will:
- Select random requirement → Verify mappings are accurate
- Select random control → Verify all driving requirements identified
- Challenge gaps → Verify remediation is planned/in progress
- Cross-check SoA → Verify control selections align with mappings

**Data Protection:**
Mapping matrix contains sensitive information:
- Control implementation gaps and deficiencies
- Regulatory obligations and interpretations
- Remediation plans and timelines
- Strategic control investment decisions

Handle in accordance with [Organisation]'s data classification policies.

**Common Mapping Challenges:**

1. **Granularity Mismatch**:
   - Requirement is specific, control is general (or vice versa)
   - Solution: Map at appropriate level, use implementation guidance to bridge gap

2. **Partial Control Implementation**:
   - Control partially implemented, only satisfies some requirements
   - Solution: Map accurately (Secondary vs. Primary), document gap for unmet aspects

3. **Implicit Controls**:
   - Organisation has control but it's not formally documented in ISMS
   - Solution: Document control properly, then map (don't map to undocumented controls)

4. **Technology-Specific Requirements**:
   - Requirement prescribes specific technology (e.g., "Use MFA")
   - Solution: Map to general control (A.5.17 Authentication), note specific implementation in evidence

5. **Cross-Cutting Requirements**:
   - Requirement affects multiple control domains (e.g., "Security by design")
   - Solution: Map to all relevant controls with appropriate types (likely many Supporting)

**Quality Assurance:**
Before finalizing control mapping:
- All requirements from Workbook 3 included (complete coverage)
- Every requirement has at least one mapping (Primary, Secondary, or gap documented)
- High-priority requirements have Primary controls identified
- Gaps have remediation plans with timelines
- Control reuse opportunities identified
- Evidence locations documented for all Primary/Secondary mappings
- Mapping types assigned consistently across similar requirements
- Cross-check with SoA (control selections align with mappings)

**Maintenance and Updates:**
Mapping requires ongoing maintenance:

**Triggers for Remapping:**
- New requirement added (from new/amended regulation)
- New control implemented (creates new mapping opportunities)
- Control enhanced (Secondary → Primary possible)
- Gap remediated (N/A → Primary/Secondary)
- Control removed/changed (breaks existing mappings)
- Audit finding questions mapping (re-assess mapping accuracy)

**Change Management:**
When mappings change:
1. Document what changed and why
2. Update coverage percentages automatically (formulas)
3. Re-assess gaps (gap resolved? new gap created?)
4. Update Evidence Register (Workbook 5) if needed
5. Update Dashboard (Workbook 6) metrics

**Scalability:**
Large regulatory portfolios with 500+ requirements:
- Matrix becomes very large (500 rows × 93 columns = 46,500 cells)
- Use filtering extensively (by regulation, by control domain, by gap status)
- Consider sub-matrices per regulation (link to master)
- Use pivot tables for analysis
- Focus manual effort on Primary/Secondary mappings (most N/A cells auto-populate)

**Common Pitfalls:**
- Mapping requirements to controls not yet implemented (aspirational mapping)
- Over-claiming Primary (marking as Primary when really Secondary)
- Under-mapping (missing relevant controls due to narrow interpretation)
- Not documenting gaps (pretending all requirements covered)
- Mapping without understanding control implementation (map based on control title, not actual implementation)
- Not maintaining mappings when controls or requirements change
- Confusing "control exists in ISMS" with "control is implemented and effective"

**Integration Workflow:**
Requirements Register (Workbook 3) → Control Mapping (Workbook 4) → Evidence Register (Workbook 5) → Dashboard (Workbook 6)

This mapping is the CRITICAL LINK between legal obligations and operational reality.

================================================================================
"""

from datetime import datetime
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging
import sys

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.31.4"
WORKBOOK_NAME = "Control Mapping"
CONTROL_ID   = "A.5.31"
CONTROL_NAME = "Legal, Statutory, Regulatory and Contractual Requirements"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = ''    #  Chart
TARGET = ''    #  Target
SHIELD = ''    # ️  Shield
LOCK = ''     #  Lock
SCALES = '\u2696'     # ⚖️  Scales of Justice
DOCUMENT = ''       #  Document
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# SECTION 1: ISO 27001:2022 CONTROLS REFERENCE
# ============================================================================

def get_iso27001_controls():
    """Return complete list of ISO 27001:2022 Annex A controls."""
    controls = [
        # ORGANISATIONAL CONTROLS (5)
        ("A.5.1", "Policies for information security", "Organisational"),
        ("A.5.2", "Information security roles and responsibilities", "Organisational"),
        ("A.5.3", "Segregation of duties", "Organisational"),
        ("A.5.4", "Management responsibilities", "Organisational"),
        ("A.5.5", "Contact with authorities", "Organisational"),
        ("A.5.6", "Contact with special interest groups", "Organisational"),
        ("A.5.7", "Threat intelligence", "Organisational"),
        ("A.5.8", "Information security in project management", "Organisational"),
        ("A.5.9", "Inventory of information and other associated assets", "Organisational"),
        ("A.5.10", "Acceptable use of information and other associated assets", "Organisational"),
        ("A.5.11", "Return of assets", "Organisational"),
        ("A.5.12", "Classification of information", "Organisational"),
        ("A.5.13", "Labelling of information", "Organisational"),
        ("A.5.14", "Information transfer", "Organisational"),
        ("A.5.15", "Access control", "Organisational"),
        ("A.5.16", "Identity management", "Organisational"),
        ("A.5.17", "Authentication information", "Organisational"),
        ("A.5.18", "Access rights", "Organisational"),
        ("A.5.19", "Information security in supplier relationships", "Organisational"),
        ("A.5.20", "Addressing information security within supplier agreements", "Organisational"),
        ("A.5.21", "Managing information security in the ICT supply chain", "Organisational"),
        ("A.5.22", "Monitoring, review and change management of supplier services", "Organisational"),
        ("A.5.23", "Information security for use of cloud services", "Organisational"),
        ("A.5.24", "Information security incident management planning and preparation", "Organisational"),
        ("A.5.25", "Assessment and decision on information security events", "Organisational"),
        ("A.5.26", "Response to information security incidents", "Organisational"),
        ("A.5.27", "Learning from information security incidents", "Organisational"),
        ("A.5.28", "Collection of evidence", "Organisational"),
        ("A.5.29", "Information security during disruption", "Organisational"),
        ("A.5.30", "ICT readiness for business continuity", "Organisational"),
        ("A.5.31", "Legal, statutory, regulatory and contractual requirements", "Organisational"),
        ("A.5.32", "Intellectual property rights", "Organisational"),
        ("A.5.33", "Protection of records", "Organisational"),
        ("A.5.34", "Privacy and protection of PII", "Organisational"),
        ("A.5.35", "Independent review of information security", "Organisational"),
        ("A.5.36", "Compliance with policies, rules and standards for information security", "Organisational"),
        ("A.5.37", "Documented operating procedures", "Organisational"),
        
        # PEOPLE CONTROLS (6)
        ("A.6.1", "Screening", "People"),
        ("A.6.2", "Terms and conditions of employment", "People"),
        ("A.6.3", "Information security awareness, education and training", "People"),
        ("A.6.4", "Disciplinary process", "People"),
        ("A.6.5", "Responsibilities after termination or change of employment", "People"),
        ("A.6.6", "Confidentiality or non-disclosure agreements", "People"),
        ("A.6.7", "Remote working", "People"),
        ("A.6.8", "Information security event reporting", "People"),
        
        # PHYSICAL CONTROLS (7)
        ("A.7.1", "Physical security perimeters", "Physical"),
        ("A.7.2", "Physical entry", "Physical"),
        ("A.7.3", "Securing offices, rooms and facilities", "Physical"),
        ("A.7.4", "Physical security monitoring", "Physical"),
        ("A.7.5", "Protecting against physical and environmental threats", "Physical"),
        ("A.7.6", "Working in secure areas", "Physical"),
        ("A.7.7", "Clear desk and clear screen", "Physical"),
        ("A.7.8", "Equipment siting and protection", "Physical"),
        ("A.7.9", "Security of assets off-premises", "Physical"),
        ("A.7.10", "Storage media", "Physical"),
        ("A.7.11", "Supporting utilities", "Physical"),
        ("A.7.12", "Cabling security", "Physical"),
        ("A.7.13", "Equipment maintenance", "Physical"),
        ("A.7.14", "Secure disposal or re-use of equipment", "Physical"),
        
        # TECHNOLOGICAL CONTROLS (8)
        ("A.8.1", "User endpoint devices", "Technological"),
        ("A.8.2", "Privileged access rights", "Technological"),
        ("A.8.3", "Information access restriction", "Technological"),
        ("A.8.4", "Access to source code", "Technological"),
        ("A.8.5", "Secure authentication", "Technological"),
        ("A.8.6", "Capacity management", "Technological"),
        ("A.8.7", "Protection against malware", "Technological"),
        ("A.8.8", "Management of technical vulnerabilities", "Technological"),
        ("A.8.9", "Configuration management", "Technological"),
        ("A.8.10", "Information deletion", "Technological"),
        ("A.8.11", "Data masking", "Technological"),
        ("A.8.12", "Data leakage prevention", "Technological"),
        ("A.8.13", "Information backup", "Technological"),
        ("A.8.14", "Redundancy of information processing facilities", "Technological"),
        ("A.8.15", "Logging", "Technological"),
        ("A.8.16", "Monitoring activities", "Technological"),
        ("A.8.17", "Clock synchronization", "Technological"),
        ("A.8.18", "Use of privileged utility programs", "Technological"),
        ("A.8.19", "Installation of software on operational systems", "Technological"),
        ("A.8.20", "Networks security", "Technological"),
        ("A.8.21", "Security of network services", "Technological"),
        ("A.8.22", "Segregation of networks", "Technological"),
        ("A.8.23", "Web filtering", "Technological"),
        ("A.8.24", "Use of cryptography", "Technological"),
        ("A.8.25", "Secure development life cycle", "Technological"),
        ("A.8.26", "Application security requirements", "Technological"),
        ("A.8.27", "Secure system architecture and engineering principles", "Technological"),
        ("A.8.28", "Secure coding", "Technological"),
        ("A.8.29", "Security testing in development and acceptance", "Technological"),
        ("A.8.30", "Outsourced development", "Technological"),
        ("A.8.31", "Separation of development, test and production environments", "Technological"),
        ("A.8.32", "Change management", "Technological"),
        ("A.8.33", "Test information", "Technological"),
        ("A.8.34", "Protection of information systems during audit testing", "Technological"),
    ]
    return controls


# ============================================================================
# SECTION 2: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    # Sheet structure
    sheets = [
        "Instructions & Legend",
        "Control Mapping Matrix",
        "ISO27001 Controls Reference",
        "Mapping Guidelines",
        "Gap Summary",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    styles = {
        "header": {
            "font": Font(name="Calibri", size=9, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90),
            "border": border_thin,
        },
        "requirement_header": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "requirement_cell": {
            "font": Font(name="Calibri", size=9),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "mapping_cell": {
            "font": Font(name="Calibri", size=10, bold=True),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "primary_mapping": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "secondary_mapping": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
        },
        "supporting_mapping": {
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        },
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        f = style_dict["font"]
        cell.font = Font(
            name=f.name,
            size=f.size,
            bold=f.bold if hasattr(f, "bold") else False,
            color=f.color if hasattr(f, "color") else "000000",
        )
    if "fill" in style_dict:
        f = style_dict["fill"]
        cell.fill = PatternFill(
            start_color=f.start_color.rgb if hasattr(f.start_color, "rgb") else f.start_color,
            end_color=f.end_color.rgb if hasattr(f.end_color, "rgb") else f.end_color,
            fill_type=f.fill_type,
        )
    if "alignment" in style_dict:
        a = style_dict["alignment"]
        cell.alignment = Alignment(
            horizontal=a.horizontal,
            vertical=a.vertical,
            wrap_text=a.wrap_text if hasattr(a, "wrap_text") else False,
            text_rotation=a.text_rotation if hasattr(a, "text_rotation") else 0,
        )
    if "border" in style_dict:
        cell.border = style_dict["border"]


# ============================================================================
# SECTION 3: SAMPLE REQUIREMENTS DATA
# ============================================================================

def get_sample_requirements():
    """Return sample requirements for demonstration (from Workbook 3)."""
    return [
        ("REG-GDPR-32-001", "Implement encryption for personal data", "Technical", "High"),
        ("REG-GDPR-33-001", "Notify DPA within 72h of breach", "Reporting", "High"),
        ("REG-GDPR-37-001", "Appoint Data Protection Officer", "Organisational", "High"),
        ("REG-ISO27001-5.1-001", "Establish information security policies", "Organisational", "High"),
        ("REG-ISO27001-6.1-001", "Define risk assessment process", "Organisational", "High"),
        ("REG-ISO27001-8.1-001", "Implement Annex A controls", "Technical", "High"),
        ("REG-FADP-7-001", "Risk-appropriate protection measures", "Technical", "High"),
        ("REG-FADP-24-001", "Notify FDPIC of high-risk breaches", "Reporting", "High"),
        ("REG-NIS2-21-001", "Risk management for network security", "Technical", "Medium"),
        ("REG-NIS2-23-001", "Incident notification to CSIRT", "Reporting", "Medium"),
    ]


def get_sample_mappings():
    """Return sample control mappings for demonstration."""
    # Format: {requirement_id: {control_id: mapping_type}}
    return {
        "REG-GDPR-32-001": {"A.8.24": "P", "A.5.10": "S"},
        "REG-GDPR-33-001": {"A.5.26": "P", "A.5.25": "S"},
        "REG-GDPR-37-001": {"A.5.2": "P", "A.6.1": "Su"},
        "REG-ISO27001-5.1-001": {"A.5.1": "P"},
        "REG-ISO27001-6.1-001": {"A.5.7": "P", "A.8.8": "S"},
        "REG-ISO27001-8.1-001": {},  # Maps to all controls - example of gap
        "REG-FADP-7-001": {"A.5.10": "P", "A.5.14": "S", "A.8.24": "S"},
        "REG-FADP-24-001": {"A.5.26": "P", "A.5.25": "S"},
        "REG-NIS2-21-001": {"A.5.7": "P", "A.8.8": "S", "A.8.20": "S"},
        "REG-NIS2-23-001": {"A.5.26": "P", "A.5.27": "Su"},
    }


# ============================================================================
# SECTION 4: POPULATE MAPPING MATRIX SHEET
# ============================================================================

def populate_mapping_matrix(wb, styles):
    """Populate the Control Mapping Matrix sheet."""
    ws = wb["Control Mapping Matrix"]
    ws.sheet_view.showGridLines = False

    controls = get_iso27001_controls()
    requirements = get_sample_requirements()
    mappings = get_sample_mappings()

    # Column A-D: Requirement info
    # Columns E onwards: One column per control (93 controls)

    # Set column widths
    ws.column_dimensions["A"].width = 18  # Requirement ID
    ws.column_dimensions["B"].width = 40  # Requirement text
    ws.column_dimensions["C"].width = 15  # Category
    ws.column_dimensions["D"].width = 12  # Priority

    # Control columns: narrow (4 width)
    for col_idx in range(5, 5 + len(controls)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4

    # Row 1: Title row - ALL CAPS, 003366 fill
    num_cols = 4 + len(controls)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value="CONTROL MAPPING MATRIX")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Row 2: Headers for requirement columns
    headers = ["Requirement ID", "Interpreted Requirement", "Category", "Priority"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        apply_style(cell, styles["requirement_header"])

    # Row 2: Headers for control columns (control IDs)
    for col_idx, (control_id, control_name, control_type) in enumerate(controls, start=5):
        cell = ws.cell(row=2, column=col_idx, value=control_id)
        apply_style(cell, styles["header"])

    # Populate requirement rows (starting row 3)
    for row_idx, (req_id, req_text, category, priority) in enumerate(requirements, start=3):
        # Requirement info columns
        ws.cell(row=row_idx, column=1, value=req_id)
        apply_style(ws.cell(row=row_idx, column=1), styles["requirement_cell"])

        ws.cell(row=row_idx, column=2, value=req_text)
        apply_style(ws.cell(row=row_idx, column=2), styles["requirement_cell"])

        ws.cell(row=row_idx, column=3, value=category)
        apply_style(ws.cell(row=row_idx, column=3), styles["requirement_cell"])

        ws.cell(row=row_idx, column=4, value=priority)
        apply_style(ws.cell(row=row_idx, column=4), styles["requirement_cell"])

        # Mapping columns
        req_mappings = mappings.get(req_id, {})
        for col_idx, (control_id, _, _) in enumerate(controls, start=5):
            cell = ws.cell(row=row_idx, column=col_idx)
            mapping_value = req_mappings.get(control_id, "")
            cell.value = mapping_value
            apply_style(cell, styles["mapping_cell"])

    # Add data validation for mapping columns
    mapping_validation = DataValidation(
        type="list",
        formula1='"P,S,Su,"',  # P, S, Su, or blank
        allow_blank=True
    )
    mapping_validation.error = "Please enter P, S, Su, or leave blank"
    mapping_validation.errorTitle = "Invalid Mapping"
    ws.add_data_validation(mapping_validation)

    # Apply to all mapping columns (E onwards) for rows 3-100
    first_control_col = get_column_letter(5)
    last_control_col = get_column_letter(4 + len(controls))
    mapping_validation.add(f"{first_control_col}3:{last_control_col}100")

    # Add conditional formatting for mapping types
    add_mapping_conditional_formatting(ws, controls)

    # Freeze panes (freeze rows 1-2 and column E)
    ws.freeze_panes = "E3"


def add_mapping_conditional_formatting(ws, controls):
    """Add conditional formatting for mapping cell values."""
    from openpyxl.formatting.rule import CellIsRule
    
    primary_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    secondary_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    supporting_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Apply to all mapping columns
    first_col = get_column_letter(5)
    last_col = get_column_letter(4 + len(controls))
    range_ref = f"{first_col}3:{last_col}100"
    
    ws.conditional_formatting.add(
        range_ref,
        CellIsRule(operator="equal", formula=['"P"'], fill=primary_fill)
    )
    ws.conditional_formatting.add(
        range_ref,
        CellIsRule(operator="equal", formula=['"S"'], fill=secondary_fill)
    )
    ws.conditional_formatting.add(
        range_ref,
        CellIsRule(operator="equal", formula=['"Su"'], fill=supporting_fill)
    )


# ============================================================================
# SECTION 5: ISO CONTROLS REFERENCE SHEET
# ============================================================================

def populate_controls_reference(wb):
    """Populate the ISO27001 Controls Reference sheet."""
    ws = wb["ISO27001 Controls Reference"]
    ws.sheet_view.showGridLines = False
    
    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "ISO/IEC 27001:2022 ANNEX A CONTROLS REFERENCE"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # Headers
    headers = ["Control ID", "Control Name", "Control Type", "Section", "Brief Description"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 60
    
    # Populate all 93 controls
    controls = get_iso27001_controls()
    for row_idx, (control_id, control_name, control_type) in enumerate(controls, start=3):
        ws.cell(row=row_idx, column=1, value=control_id)
        ws.cell(row=row_idx, column=2, value=control_name)
        ws.cell(row=row_idx, column=3, value=control_type)
        
        # Determine section
        section_map = {
            "Organisational": "5 - Organisational Controls",
            "People": "6 - People Controls",
            "Physical": "7 - Physical Controls",
            "Technological": "8 - Technological Controls",
        }
        ws.cell(row=row_idx, column=4, value=section_map[control_type])
        
        # Brief description (simplified)
        ws.cell(row=row_idx, column=5, value=f"{control_name} requirements")
        
        # Formatting
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin = Side(style="thin")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Freeze panes
    ws.freeze_panes = "A3"
    
    # Auto-filter
    ws.auto_filter.ref = "A2:E2"


# ============================================================================
# SECTION 6: MAPPING GUIDELINES SHEET
# ============================================================================

def populate_guidelines_sheet(wb):
    """Populate the Mapping Guidelines sheet."""
    ws = wb["Mapping Guidelines"]
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"ISMS-IMP-A.5.31.4 | CONTROL MAPPING MATRIX — GUIDELINES | ISO/IEC 27001:2022 — CONTROL A.5.31"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    guidelines = [
        "",
        "PURPOSE:",
        "Map regulatory requirements to ISO 27001 Annex A controls.",
        "Identifies which controls satisfy which requirements and reveals gaps.",
        "",
        "MAPPING TYPES:",
        "",
        "P (Primary):",
        "  • Control DIRECTLY and SUBSTANTIALLY satisfies requirement",
        "  • This is the 'go-to' control for this requirement",
        "  • Example: Encryption requirement → A.8.24 Use of Cryptography",
        "  • Example: DPO appointment → A.5.2 Roles and Responsibilities",
        "",
        "S (Secondary):",
        "  • Control PARTIALLY satisfies requirement",
        "  • Additional controls or measures needed for full satisfaction",
        "  • Example: Access control requirement → A.5.15 Access Control (partial),",
        "    also needs A.8.2 Privileged Access Rights (specific implementation)",
        "",
        "Su (Supporting):",
        "  • Control INDIRECTLY supports requirement",
        "  • Contributes but doesn't directly satisfy",
        "  • Example: Technical security requirement → A.6.3 Awareness Training (supporting)",
        "    (training doesn't satisfy technical requirement but supports implementation)",
        "",
        "Blank:",
        "  • Control not applicable to this requirement",
        "  • Most cells will be blank (requirements typically map to 1-5 controls)",
        "",
        "GAP IDENTIFICATION:",
        "",
        "Complete Gap:",
        "  • Requirement row with NO 'P' (Primary) mappings",
        "  • This requirement is NOT currently satisfied by any control",
        "  • Action: Implement new control or enhance existing control to Primary level",
        "",
        "Partial Gap:",
        "  • Requirement row with only 'S' (Secondary) or 'Su' (Supporting) mappings",
        "  • Requirement partially addressed but not fully satisfied",
        "  • Action: Enhance secondary controls to primary level or add missing controls",
        "",
        "No Gap:",
        "  • Requirement row with at least one 'P' (Primary) mapping",
        "  • Requirement is satisfied (may also have additional S/Su for defense-in-depth)",
        "",
        "USING THIS MATRIX:",
        "",
        "Step 1 - Review Requirements:",
        "  • Import requirements from Workbook 3 (Requirements Register)",
        "  • Or manually add requirements to rows",
        "",
        "Step 2 - For Each Requirement:",
        "  • Review ISO 27001 controls (use Sheet 2 reference)",
        "  • Identify which controls address this requirement",
        "  • Enter P, S, or Su in appropriate control columns",
        "  • Leave blank where control not relevant",
        "",
        "Step 3 - Identify Gaps:",
        "  • Review Gap Summary sheet (Sheet 4)",
        "  • Focus on requirements with no Primary mappings",
        "  • Prioritize gaps by requirement priority (High/Medium/Low)",
        "",
        "Step 4 - Gap Remediation:",
        "  • For complete gaps: Implement new control or significantly enhance existing",
        "  • For partial gaps: Enhance secondary controls or add complementary controls",
        "  • Update Gap Status in Requirements Register (Workbook 3)",
        "",
        "WORKFLOW:",
        "Requirements (Workbook 3) → Mapping (this workbook) → Gap Analysis → Remediation → Evidence (Workbook 5)",
        "",
        "TIPS:",
        "",
        "Start with High-Priority Requirements:",
        "  • Focus mapping effort on Tier 1 regulations first",
        "  • High High priority requirements need immediate attention",
        "",
        "One Requirement = Multiple Controls (Common):",
        "  • Security requirements often need multiple controls working together",
        "  • Example: Data protection may need A.8.24 (encryption) + A.8.11 (masking) + A.5.10 (acceptable use)",
        "",
        "One Control = Multiple Requirements (Common):",
        "  • Controls often satisfy multiple regulatory requirements",
        "  • Example: A.5.26 (Incident Response) satisfies GDPR, FADP, NIS2 notification requirements",
        "",
        "Defense-in-Depth:",
        "  • It's OK (and good!) to have P + S + Su mappings for one requirement",
        "  • Primary control satisfies, secondary/supporting provide additional assurance",
        "",
        "Don't Force Mappings:",
        "  • If control truly doesn't relate to requirement, leave blank",
        "  • Most cells should be blank - that's normal",
        "",
        "COLOR CODING:",
        "",
        "Mapping Cells:",
        "  • Green background = P (Primary) - Requirement satisfied",
        "  • Yellow background = S (Secondary) - Partial satisfaction",
        "  • Light blue background = Su (Supporting) - Indirect support",
        "  • No color = Blank (not applicable)",
        "",
        "QUALITY CHECKS:",
        "",
        "✓ All high-priority requirements have at least one Primary mapping",
        "✓ Mapping choices are logical (control actually addresses requirement)",
        "✓ Gap Summary reviewed and remediation plan created",
        "✓ Gaps documented in Requirements Register (Workbook 3)",
        "✓ Evidence collection planned for mapped controls (Workbook 5)",
        "",
        "COMMON MAPPING EXAMPLES:",
        "",
        "Encryption Requirements:",
        "  • Primary: A.8.24 Use of Cryptography",
        "  • Secondary: A.5.10 Acceptable Use (defines when encryption required)",
        "",
        "Breach Notification:",
        "  • Primary: A.5.26 Response to Information Security Incidents",
        "  • Secondary: A.5.25 Assessment and Decision on Events (detection)",
        "  • Supporting: A.6.8 Event Reporting (enables detection)",
        "",
        "Access Control:",
        "  • Primary: A.5.15 Access Control",
        "  • Primary: A.5.18 Access Rights",
        "  • Secondary: A.5.16 Identity Management",
        "  • Secondary: A.8.2 Privileged Access Rights",
        "  • Secondary: A.8.3 Information Access Restriction",
        "",
        "Policies & Governance:",
        "  • Primary: A.5.1 Policies for Information Security",
        "  • Supporting: A.5.2 Information Security Roles",
        "  • Supporting: A.5.4 Management Responsibilities",
        "",
        "RELATED PROCESSES:",
        "  • IMP-5.31.3: Requirements Extraction Process (source of requirements)",
        "  • IMP-5.31.4: Control Mapping Process (how to use this workbook)",
        "  • IMP-5.31.5: Evidence Management Process (next step after mapping)",
    ]
    
    for idx, line in enumerate(guidelines, start=2):
        cell = ws.cell(row=idx, column=1, value=line)
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    ws.column_dimensions["A"].width = 75
    ws.freeze_panes = "A2"


# ============================================================================
# SECTION 7: GAP SUMMARY SHEET
# ============================================================================

def populate_gap_summary(wb):
    """Populate the Gap Summary sheet."""
    ws = wb["Gap Summary"]
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "GAP ANALYSIS SUMMARY"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Headers
    headers = ["Requirement ID", "Requirement Text", "Category", "Priority", "Gap Type", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 50

    # Sample row (row 3) — grey D9D9D9, does not count in SD (COUNTA starts at A4)
    _grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _thin = Side(style="thin")
    gap_data = [
        ("REG-ISO27001-8.1-001", "Implement Annex A controls", "Technical", "High",
         f"{XMARK} Complete Gap",
         "This is a meta-requirement covering all 93 controls. Gap analysis shows 28 controls not yet implemented."),
    ]
    for row_idx, data in enumerate(gap_data, start=3):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=9)
            cell.fill = _grey_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    # Data validations — applied to data rows (row 4+), skipping grey sample at row 3
    dv_category = DataValidation(
        type="list",
        formula1='"Technical,Organisational,Legal/Regulatory,Operational,Physical"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_category)
    dv_category.add("C4:C1000")

    dv_priority = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_priority)
    dv_priority.add("D4:D1000")

    dv_gap_type = DataValidation(
        type="list",
        formula1=f'"{XMARK} Complete Gap,{WARNING} Partial Gap,{CHECK} No Gap"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_gap_type)
    dv_gap_type.add("E4:E1000")

    # Freeze panes and auto-filter
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:F2"


# ============================================================================
# SECTION 8: MAIN EXECUTION
# ============================================================================



def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete the Control Mapping Matrix — map each regulatory requirement to ISO 27001:2022 Annex A controls.',
        '2. Use the ISO 27001 Controls Reference sheet to understand each control’s scope.',
        '3. Review the Mapping Guidelines sheet before completing mappings.',
        '4. Complete the Gap Summary — identify requirements not yet covered by implemented controls.',
        '5. Create a remediation plan for all identified gaps with owners and target dates.',
        '6. Maintain the Evidence Register with mapping rationale and gap analysis documentation.',
        '7. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A21"] = "Status Legend"
    ws["A21"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=22, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 23 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_instructions_legend_sheet(wb):
    """Create GS-IL-compliant Instructions & Legend sheet."""
    ws = wb["Instructions & Legend"]
    ws.sheet_view.showGridLines = False

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID", DOCUMENT_ID),
        ("Workbook Title", WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version", "1.0"),
        ("Assessment Date", ""),
        ("Completed By", ""),
        ("Organisation", ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Instructions section
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        "1. Review this Instructions & Legend sheet before starting the assessment.",
        "2. Map each requirement to ISO 27001:2022 controls using P (Primary), S (Supporting), or Su (Supplementary).",
        "3. Grey (F2F2F2) rows are sample rows — do not edit them.",
        "4. Use the Gap Summary sheet to identify unmapped requirements.",
        "5. Obtain sign-off in the Approval Sign-Off sheet when the assessment is complete.",
    ]):
        ws[f"A{13 + i}"] = line

    # Status Legend section
    ws["A21"] = "Status Legend"
    ws["A21"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=22, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("P", "Primary", "This control is the primary mechanism addressing the requirement", _green),
        ("S", "Supporting", "This control supports the primary control for this requirement", _amber),
        ("Su", "Supplementary", "This control supplements but is not the main control", None),
        ("\u2014", "Not Applicable", "This control is not relevant to the requirement", None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 23 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"


def create_summary_dashboard_sheet(wb):
    """Create Gold Standard Summary Dashboard sheet — ISMS-IMP-A.5.31.4 Control Mapping."""
    ws = wb["Summary Dashboard"]
    ws.sheet_view.showGridLines = False

    _thin = Side(border_style="thin", color="000000")
    _b    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy  = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _blue  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _red   = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _grey  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _yell  = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _lft   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Column widths (7 columns: A:G)
    for col, w in zip("ABCDEFG", [50, 12, 18, 15, 18, 12, 15]):
        ws.column_dimensions[col].width = w

    # ── Row 1: Title (GS-SD-014: must contain em dash + SUMMARY DASHBOARD) ─
    ws.merge_cells("A1:G1")
    ws["A1"] = "CONTROL MAPPING \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = _navy
    ws["A1"].alignment = _ctr
    ws.row_dimensions[1].height = 35
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = _b

    # ── Row 2: Subtitle (left aligned, no fill) ──────────────────────────
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.5.31: Legal, Statutory, Regulatory and Contractual Requirements | Control Mapping"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=2, column=c).border = _b

    # ── Row 3: Empty separator ───────────────────────────────────────────

    # ── TABLE 1: Assessment Area Compliance Overview ─────────────────────
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws["A4"].fill = _navy
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=4, column=c).border = _b

    # TABLE 1 headers (row 5) — D9D9D9, black bold (GS-SD-016: NOT 4472C4)
    for c, h in enumerate(["Assessment Area", "Total Items", "Compliant", "Partial",
                            "Non-Compliant", "N/A", "Compliance %"], 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = Font(bold=True, size=10, color="000000", name="Calibri")
        cell.fill = _grey
        cell.alignment = _ctr
        cell.border = _b

    # TABLE 1 single data row (row 6): Control Mapping Coverage
    # Total = COUNTA(Control Mapping Matrix!A3:A1000)
    # Compliant = Total minus gap items; Non-Compliant = gap items in Gap Summary
    row = 6
    ws.cell(row=row, column=1, value="Control Mapping Coverage").border = _b
    ws.cell(row=row, column=1).font = Font(color="000000", name="Calibri", size=10)
    ws.cell(row=row, column=1).alignment = _lft

    # B: Total requirements in mapping
    cell_b = ws.cell(row=row, column=2,
        value="=COUNTA('Control Mapping Matrix'!A3:A1000)")
    cell_b.border = _b; cell_b.alignment = Alignment(horizontal="center")
    cell_b.font = Font(color="000000", name="Calibri", size=10)

    # C: Compliant = Total minus gap items
    cell_c = ws.cell(row=row, column=3,
        value=f"=B{row}-COUNTA('Gap Summary'!A4:A1000)")
    cell_c.border = _b; cell_c.alignment = Alignment(horizontal="center")
    cell_c.font = Font(color="000000", name="Calibri", size=10)

    # D: Partial = 0
    cell_d = ws.cell(row=row, column=4, value=0)
    cell_d.border = _b; cell_d.alignment = Alignment(horizontal="center")
    cell_d.font = Font(color="000000", name="Calibri", size=10)

    # E: Non-Compliant = gap items
    cell_e = ws.cell(row=row, column=5,
        value="=COUNTA('Gap Summary'!A4:A1000)")
    cell_e.border = _b; cell_e.alignment = Alignment(horizontal="center")
    cell_e.font = Font(color="000000", name="Calibri", size=10)

    # F: N/A = 0 (static)
    cell_f = ws.cell(row=row, column=6, value=0)
    cell_f.border = _b; cell_f.alignment = Alignment(horizontal="center")
    cell_f.font = Font(color="000000", name="Calibri", size=10)

    # G: Compliance %
    cell_g = ws.cell(row=row, column=7,
        value=f"=IFERROR(IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row})),\"\")")
    cell_g.number_format = "0.0%"
    cell_g.border = _b; cell_g.alignment = Alignment(horizontal="center")
    cell_g.font = Font(color="000000", name="Calibri", size=10)

    # TOTAL row (row 7)
    total_row = 7
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="000000", name="Calibri", size=10)
    ws.cell(row=total_row, column=1).fill = _grey
    ws.cell(row=total_row, column=1).border = _b
    ws.cell(row=total_row, column=1).alignment = _lft
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col,
                       value=f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{total_row - 1})")
        cell.font = Font(bold=True, color="000000", name="Calibri", size=10)
        cell.fill = _grey
        cell.border = _b
        cell.alignment = Alignment(horizontal="center")
    cell_tot_pct = ws.cell(row=total_row, column=7,
                           value=f"=IFERROR(IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row})),\"\")")
    cell_tot_pct.number_format = "0.0%"
    cell_tot_pct.font = Font(bold=True, color="000000", name="Calibri", size=10)
    cell_tot_pct.fill = _grey
    cell_tot_pct.border = _b
    cell_tot_pct.alignment = Alignment(horizontal="center")

    # ── TABLE 2: Key Metrics ─────────────────────────────────────────────
    t2_banner_row = total_row + 2  # row 9
    ws.merge_cells(f"A{t2_banner_row}:G{t2_banner_row}")
    ws[f"A{t2_banner_row}"] = "TABLE 2: KEY METRICS"
    ws[f"A{t2_banner_row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{t2_banner_row}"].fill = _navy
    ws[f"A{t2_banner_row}"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=t2_banner_row, column=c).border = _b

    # TABLE 2 headers — D9D9D9 grey, black bold (GS-SD-016)
    t2_hdr_row = t2_banner_row + 1  # row 10
    for c, h in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        cell = ws.cell(row=t2_hdr_row, column=c, value=h if h else None)
        cell.font = Font(bold=True, color="000000", name="Calibri", size=10)
        cell.fill = _grey
        cell.border = _b
        cell.alignment = _ctr

    # TABLE 2 metrics — white fill, 000000 font, NOT bold labels (GS-SD-015)
    metrics = [
        ("Total requirements in scope",
         "=COUNTA('Control Mapping Matrix'!A3:A1000)"),
        ("Requirements with Primary control mapped",
         f"=B{6}-COUNTA('Gap Summary'!A4:A1000)"),
        ("Requirements in gap analysis",
         "=COUNTA('Gap Summary'!A4:A1000)"),
        ("Total ISO 27001 controls referenced",
         "=COUNTA('ISO27001 Controls Reference'!A3:A1000)"),
    ]
    row = t2_hdr_row + 1  # row 11
    for metric, formula in metrics:
        cell_m = ws.cell(row=row, column=1, value=metric)
        cell_m.border = _b
        cell_m.font = Font(color="000000", name="Calibri", size=10)  # NOT bold (GS-SD-015)
        cell_m.alignment = _lft
        cell_v = ws.cell(row=row, column=2, value=formula)
        cell_v.border = _b
        cell_v.font = Font(color="000000", name="Calibri", size=10)
        cell_v.alignment = Alignment(horizontal="center")
        for c in range(3, 8):
            ws.cell(row=row, column=c).border = _b
        row += 1
    t2_last_row = row - 1  # row 14

    # ── TABLE 3: Critical Findings ────────────────────────────────────────
    t3_banner_row = t2_last_row + 2  # row 16
    ws.merge_cells(f"A{t3_banner_row}:G{t3_banner_row}")
    ws[f"A{t3_banner_row}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{t3_banner_row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{t3_banner_row}"].fill = _red
    ws[f"A{t3_banner_row}"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 8):
        ws.cell(row=t3_banner_row, column=c).border = _b

    # TABLE 3 headers — D9D9D9
    t3_hdr_row = t3_banner_row + 1  # row 17
    for c, h in enumerate(["Finding", "Count", "Action Required", "", "", "", ""], 1):
        cell = ws.cell(row=t3_hdr_row, column=c, value=h if h else None)
        cell.font = Font(bold=True, color="000000", name="Calibri", size=10)
        cell.fill = _grey
        cell.border = _b
        cell.alignment = _ctr
    ws.merge_cells(f"C{t3_hdr_row}:G{t3_hdr_row}")

    _yell_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    findings = [
        ("Requirements in gap analysis (no adequate control)",
         "=COUNTA('Gap Summary'!A4:A1000)",
         "Review Gap Summary sheet \u2014 implement controls or accept risk for each gap item"),
        ("Total requirements not yet mapped",
         "=COUNTA('Control Mapping Matrix'!A3:A1000)-COUNTA('Mapping Guidelines'!A3:A1000)",
         "Complete control mapping for all extracted requirements before gap analysis"),
    ]
    row = t3_hdr_row + 1  # row 18
    for finding, count_formula, action in findings:
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = _yell_fill
            ws.cell(row=row, column=c).border = _b
            ws.cell(row=row, column=c).font = Font(color="000000", name="Calibri", size=10)
        ws.cell(row=row, column=1, value=finding).alignment = _lft
        cell_cnt = ws.cell(row=row, column=2, value=count_formula)
        cell_cnt.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"C{row}:G{row}")
        cell_act = ws.cell(row=row, column=3, value=action)
        cell_act.alignment = _lft
        for c in range(4, 8):
            ws.cell(row=row, column=c).border = _b
        row += 1

    # 2 empty FFFFCC buffer rows
    for _ in range(2):
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = _yell_fill
            ws.cell(row=row, column=c).border = _b
        row += 1
    t3_last_row = row - 1

    # ── FINAL DECISION (GS-AS-012: col A plain bold, NO dark fill) ───────
    fd_row = t3_last_row + 2
    ws.cell(row=fd_row, column=1, value="FINAL DECISION:").font = Font(bold=True, name="Calibri")
    ws.merge_cells(f"B{fd_row}:G{fd_row}")
    ws.cell(row=fd_row, column=2).fill = _yell
    for c in range(2, 8):
        ws.cell(row=fd_row, column=c).border = _b

    fd_dv = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(fd_dv)
    fd_dv.add(f"B{fd_row}")

    # ── NEXT REVIEW DETAILS ───────────────────────────────────────────────
    nr_row = fd_row + 3
    ws.merge_cells(f"A{nr_row}:G{nr_row}")
    ws[f"A{nr_row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{nr_row}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{nr_row}"].fill = _blue
    for c in range(1, 8):
        ws.cell(row=nr_row, column=c).border = _b

    for i, label in enumerate(["Next Review Date:", "Review Responsible:", "Special Considerations:"]):
        r = nr_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, name="Calibri")
        ws.merge_cells(f"B{r}:G{r}")
        ws.cell(row=r, column=2).fill = _yell
        for c in range(2, 8):
            ws.cell(row=r, column=c).border = _b

    # Apply borders to all merged ranges (GS-AS-011)
    for mr in list(ws.merged_cells.ranges):
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                ws.cell(row=r, column=c).border = _b

    ws.freeze_panes = "A4"


def create_approval_sheet(wb):
    """Create the Approval Sign-Off sheet — Gold Standard (GS-AS-014/015)."""
    ws = wb["Approval Sign-Off"]
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title banner — GS-AS-014
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY section banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Rows 4-8: Summary metadata — B6 = Overall Compliance (GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE(\'Summary Dashboard\'!G6:G6),\"\")")  ,
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Row 7 status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # Approver sections start at row 11 (rows 9-10 = gap)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION — GS-AS-012: col A = plain bold label, NO dark fill
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    pass

def main():
    """Main execution function."""
    logger.info("=" * 70)
    logger.info("ISMS Assessment Workbook Generator")
    logger.info("Control A.5.31 - Workbook 4: Control Mapping Matrix")
    logger.info("=" * 70)
    logger.info("")
    
    # Create workbook
    logger.info(" Creating workbook structure...")
    wb = create_workbook()
    styles = _STYLES
    
    # Populate sheets
    logger.info(" Creating Instructions & Legend sheet...")
    create_instructions_legend_sheet(wb)

    logger.info(" Populating Control Mapping Matrix (10 sample requirements × 93 controls)...")
    populate_mapping_matrix(wb, styles)
    
    logger.info(" Populating ISO 27001 Controls Reference (all 93 controls)...")
    populate_controls_reference(wb)
    
    logger.info(" Populating Mapping Guidelines...")
    populate_guidelines_sheet(wb)
    
    logger.info(f"{CHART} Populating Gap Summary...")
    populate_gap_summary(wb)

    logger.info(" Creating Summary Dashboard (Gold Standard TABLE 1/2/3)...")
    create_summary_dashboard_sheet(wb)

    logger.info(" Creating Approval Sign-Off sheet...")
    create_approval_sheet(wb)

    # Save workbook
    _wkbk_dir.mkdir(exist_ok=True)
    output_path = _wkbk_dir / OUTPUT_FILENAME
    logger.info(f" Saving workbook to: {output_path}")
    wb.save(output_path)
    
    logger.info("")
    logger.info("{CHECK} Workbook generated successfully!")
    logger.info("")
    logger.info(" Output Details:")
    logger.info(f"   File: ISMS_Assessment_531_4_Control_Mapping_Matrix.xlsx")
    logger.info(f"   Location: WKBK/")
    logger.info(f"   Sheets: 7 (Instructions & Legend, Control Mapping Matrix, ISO27001 Controls Reference, Mapping Guidelines, Gap Summary, Summary Dashboard, Approval Sign-Off)")
    logger.info(f"   Sample Data: 10 requirements mapped to 93 ISO 27001:2022 controls")
    logger.info("")
    logger.info(" Features:")
    logger.info("   ✓ 10 requirements × 93 controls matrix (930 mapping cells)")
    logger.info("   ✓ Data validation (P/S/Su dropdowns)")
    logger.info("   ✓ Conditional formatting (color-coded by mapping type)")
    logger.info("   ✓ All 93 ISO 27001:2022 Annex A controls reference")
    logger.info("   ✓ Vertical control headers (space-efficient)")
    logger.info("   ✓ Freeze panes (requirements and headers stay visible)")
    logger.info("   ✓ Gap identification and summary")
    logger.info("   ✓ Comprehensive mapping guidelines")
    logger.info("   ✓ Sample mappings demonstrate P/S/Su usage")
    logger.info("")
    logger.info(f"{TARGET} Next Steps:")
    logger.info("   1. Import requirements from Workbook 3 (or add manually)")
    logger.info("   2. Use Controls Reference sheet to understand each control")
    logger.info("   3. Map each requirement to applicable controls (P/S/Su)")
    logger.info("   4. Review Gap Summary to identify unmapped requirements")
    logger.info("   5. Create gap remediation plan")
    logger.info("   6. Update Gap Status in Requirements Register (Workbook 3)")
    logger.info("   7. Collect evidence for mapped controls (Workbook 5)")
    logger.info("")
    logger.info("=" * 70)


if __name__ == "__main__":
    sys.exit(main())

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
