#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.17.1 - Authentication Policy and Standards Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.17: Authentication Information
Assessment Domain 1 of 3: Authentication Policy and Standards

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific authentication information management infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Authentication mechanism types and security requirements (match your systems)
2. Password/passphrase complexity and rotation policies (adapt to your standards)
3. Credential lifecycle stages and responsible owner roles
4. MFA applicability criteria and supported methods
5. Privileged credential management and vault integration scope

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.17 Authentication Information Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
authentication information management controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Authentication Policy and Standards under ISO 27001:2022 Control A.5.17. Supports evidence-based evaluation of authentication policy adherence, credential lifecycle compliance, and system security configuration.

**Assessment Scope:**
- Authentication policy definition and coverage completeness
- Credential lifecycle management process compliance
- Password system configuration against policy standards
- MFA implementation scope and coverage tracking
- Privileged credential management and storage controls
- User acknowledgment and awareness of authentication requirements
- Evidence collection for access management and compliance audits

**Generated Workbook Structure:**
1. Password Policy
2. MFA Requirements
3. Credential Standards
4. User Responsibilities
5. System Requirements
6. Evidence Register
7. Approval Sign-Off
8. Summary Dashboard

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 3 domains covering Authentication Information controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a517_1_authentication_policy_standards.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a517_1_authentication_policy_standards.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a517_1_authentication_policy_standards.py --date 20250115

Output:
    File: ISMS-IMP-A.5.17.1_Authentication_Policy_and_Standards_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.17
Assessment Domain:    1 of 3 (Authentication Policy and Standards)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.17: Authentication Information Policy (Governance)
    - ISMS-IMP-A.5.17.1: Authentication Policy and Standards (Domain 1)
    - ISMS-IMP-A.5.17.2: Credential Lifecycle Management (Domain 2)
    - ISMS-IMP-A.5.17.3: Password System Assessment (Domain 3)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.17.1 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive authentication information management details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review authentication policies and credential management procedures annually or when authentication standards evolve, new systems are introduced, or security incidents reveal gaps.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

import logging
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.17.1"
WORKBOOK_NAME = "Authentication Policy and Standards"
CONTROL_ID = "A.5.17"
CONTROL_NAME = "Authentication Information"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
GREY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP_LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass


def set_column_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_header_row(ws, row: int, headers: list, fill=SUBHEADER_FILL, font=SUBHEADER_FONT):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def add_data_validation(ws, cell_range: str, formula: str):
    dv = DataValidation(type="list", formula1=formula, showDropDown=False, allowBlank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)



def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete Password Policy — assess password length, complexity, history, and expiry requirements.',
        '2. Complete MFA Requirements — document MFA mandate per system tier and user type.',
        '3. Complete Credential Standards — review credential standards for service accounts and APIs.',
        '4. Complete User Responsibilities — assess user awareness of credential protection obligations.',
        '5. Complete System Requirements — evaluate system enforcement of authentication controls.',
        '6. Maintain the Evidence Register with policy documentation and system configuration exports.',
        '7. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A21"] = "Status Legend"
    ws["A21"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=22, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 23 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_password_policy_sheet(wb: Workbook):
    ws = wb.create_sheet("Password Policy")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "PASSWORD POLICY REQUIREMENTS"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"].value = f"{CONTROL_REF} | Define and enforce password complexity and lifecycle requirements"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = CENTER_ALIGN

    headers = ["Requirement Category", "Policy Requirement", "Standard Value", "Current Setting", "Compliant", "Implementation Notes", "Exception Process"]
    create_header_row(ws, 3, headers)

    requirements = [
        ("Minimum Length", "Minimum password length", ">=12 characters", "", "", "", ""),
        ("Maximum Length", "Maximum password length (no artificial limit)", ">=128 characters", "", "", "", ""),
        ("Complexity - Uppercase", "Require uppercase letters", "Yes", "", "", "", ""),
        ("Complexity - Lowercase", "Require lowercase letters", "Yes", "", "", "", ""),
        ("Complexity - Numbers", "Require numeric characters", "Yes", "", "", "", ""),
        ("Complexity - Special", "Require special characters", "Yes", "", "", "", ""),
        ("Password History", "Number of previous passwords prevented", ">=12 passwords", "", "", "", ""),
        ("Maximum Age", "Maximum password age before mandatory change", "<=365 days", "", "", "", ""),
        ("Minimum Age", "Minimum age before password can be changed", ">=1 day", "", "", "", ""),
        ("Initial Password", "Temporary password requirements", "Force change on first login", "", "", "", ""),
        ("Common Passwords", "Block commonly used/breached passwords", "Yes - use blocklist", "", "", "", ""),
        ("Sequential Characters", "Block sequential characters (abc, 123)", "Yes", "", "", "", ""),
        ("Username in Password", "Block username or variations in password", "Yes", "", "", "", ""),
        ("Dictionary Words", "Block dictionary words", "Recommended", "", "", "", ""),
        ("Password Hints", "Disable password hints", "Yes", "", "", "", ""),
        ("Account Lockout Threshold", "Failed attempts before lockout", "5-10 attempts", "", "", "", ""),
        ("Lockout Duration", "Account lockout duration", ">=15 minutes", "", "", "", ""),
        ("Password Storage", "Secure storage algorithm", "bcrypt/Argon2/PBKDF2", "", "", "", ""),
        ("Transmission Security", "Password transmission encryption", "TLS 1.2+ required", "", "", "", ""),
    ]

    row = 4
    for req in requirements:
        for col, value in enumerate(req, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [4, 5, 6, 7]:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"E4:E{row+5}", '"Yes,No,Partial,N/A"')

    set_column_widths(ws, {"A": 22, "B": 40, "C": 22, "D": 20, "E": 12, "F": 30, "G": 25})
    ws.freeze_panes = "A4"
    logger.info("Created Password Policy sheet")


def create_mfa_requirements_sheet(wb: Workbook):
    ws = wb.create_sheet("MFA Requirements")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "MULTI-FACTOR AUTHENTICATION REQUIREMENTS"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"{CONTROL_REF} | Define MFA requirements by system and access type"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = CENTER_ALIGN

    headers = ["System/Access Type", "MFA Required", "Approved Methods", "Fallback Method", "Current Status", "Implementation Date", "Owner", "Notes"]
    create_header_row(ws, 3, headers)

    systems = [
        ("VPN Remote Access", "MANDATORY", "TOTP, Push Notification, FIDO2", "SMS (emergency only)", "", "", "", ""),
        ("Cloud Applications (M365, etc.)", "MANDATORY", "Authenticator App, FIDO2", "Phone call", "", "", "", ""),
        ("Privileged Access (Admin)", "MANDATORY", "Hardware Token, FIDO2", "Authenticator App", "", "", "", ""),
        ("Email (External Access)", "MANDATORY", "TOTP, Push Notification", "SMS", "", "", "", ""),
        ("Internal Applications (Sensitive)", "MANDATORY", "Authenticator App, FIDO2", "SMS", "", "", "", ""),
        ("Internal Applications (Standard)", "RECOMMENDED", "Authenticator App", "N/A", "", "", "", ""),
        ("Database Access (Production)", "MANDATORY", "Hardware Token, Certificate", "TOTP", "", "", "", ""),
        ("Source Code Repository", "MANDATORY", "TOTP, FIDO2", "Authenticator App", "", "", "", ""),
        ("CI/CD Pipeline", "MANDATORY", "Service Account + Certificate", "Hardware Token", "", "", "", ""),
        ("Customer Portal (Admin)", "MANDATORY", "TOTP, Push Notification", "SMS", "", "", "", ""),
        ("Customer Portal (Users)", "RECOMMENDED", "TOTP, Email OTP", "Security Questions", "", "", "", ""),
        ("Workstation Login", "CONDITIONAL", "Windows Hello, Smart Card", "Password only", "", "", "", ""),
        ("Server Console Access", "MANDATORY", "Hardware Token, Certificate", "TOTP", "", "", "", ""),
        ("Cloud Infrastructure (AWS/Azure)", "MANDATORY", "FIDO2, Hardware Token", "Authenticator App", "", "", "", ""),
    ]

    row = 4
    for system in systems:
        for col, value in enumerate(system, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [5, 6, 7, 8]:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"B4:B{row+5}", '"MANDATORY,RECOMMENDED,CONDITIONAL,NOT REQUIRED"')
    add_data_validation(ws, f"E4:E{row+5}", '"Implemented,In Progress,Planned,Not Started,Exception"')

    set_column_widths(ws, {"A": 30, "B": 14, "C": 32, "D": 22, "E": 16, "F": 18, "G": 18, "H": 25})
    ws.freeze_panes = "A4"
    logger.info("Created MFA Requirements sheet")


def create_credential_standards_sheet(wb: Workbook):
    ws = wb.create_sheet("Credential Standards")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "CREDENTIAL TYPE STANDARDS"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"{CONTROL_REF} | Document security requirements per credential type"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = CENTER_ALIGN

    headers = ["Credential Type", "Use Cases", "Security Requirements", "Storage Requirements", "Rotation Policy", "Revocation Process", "Approved Products", "Notes"]
    create_header_row(ws, 3, headers)

    credentials = [
        ("User Passwords", "End-user authentication", "Min 12 chars, complexity", "bcrypt/Argon2, salted", "Annual or incident", "Immediate on termination", "N/A", ""),
        ("Service Account Passwords", "Application-to-application", "Min 24 chars, random", "Secrets manager (Vault)", "Quarterly", "Change on compromise", "HashiCorp Vault, Azure Key Vault", ""),
        ("API Keys", "API authentication", "Min 32 chars, random", "Secrets manager", "Quarterly", "Immediate revocation", "HashiCorp Vault, AWS Secrets Manager", ""),
        ("SSH Keys", "Server/system access", "RSA 4096-bit or Ed25519", "Encrypted key store", "Annual", "Remove from authorised_keys", "OpenSSH", ""),
        ("SSL/TLS Certificates", "Server authentication", "RSA 2048+ or ECC P-256+", "HSM or secure store", "Annual (max 1 year)", "CRL/OCSP update", "DigiCert, Let's Encrypt", ""),
        ("Client Certificates", "User/device authentication", "RSA 2048+ or ECC P-256+", "Smart card or TPM", "Annual", "CRL/OCSP update", "Internal PKI", ""),
        ("TOTP Seeds", "MFA time-based codes", "Min 160-bit secret", "Encrypted in auth system", "On device change", "Remove from account", "Microsoft Authenticator, Google Auth", ""),
        ("FIDO2/WebAuthn", "Passwordless MFA", "Platform authenticator", "Device TPM/Secure Element", "N/A (hardware bound)", "Remove registration", "YubiKey, Windows Hello", ""),
        ("Hardware Tokens", "High-security MFA", "FIPS 140-2 Level 2+", "Physical custody", "5 years or damage", "Return and deactivate", "YubiKey, RSA SecurID", ""),
        ("Biometric Data", "Local device auth", "On-device processing only", "Never stored centrally", "N/A", "Device wipe", "Windows Hello, Touch ID", ""),
        ("Recovery Codes", "MFA backup", "Random 8-digit codes", "User-managed secure storage", "On use or annually", "Regenerate", "N/A", ""),
    ]

    row = 4
    for cred in credentials:
        for col, value in enumerate(cred, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col == 8:
                cell.fill = INPUT_FILL
        row += 1

    set_column_widths(ws, {"A": 22, "B": 25, "C": 28, "D": 28, "E": 20, "F": 25, "G": 30, "H": 25})
    ws.freeze_panes = "A4"
    logger.info("Created Credential Standards sheet")


def create_user_responsibilities_sheet(wb: Workbook):
    ws = wb.create_sheet("User Responsibilities")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "USER RESPONSIBILITIES FOR AUTHENTICATION INFORMATION"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    ws["A2"].value = f"{CONTROL_REF} | Define user obligations for handling authentication credentials"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = CENTER_ALIGN

    headers = ["Responsibility", "Description", "Requirement Level", "Enforcement Method", "Training Required", "Notes"]
    create_header_row(ws, 3, headers)

    responsibilities = [
        ("Password Confidentiality", "Never share passwords with anyone including IT staff", "MANDATORY", "Policy acknowledgment, monitoring", "Yes", ""),
        ("Unique Passwords", "Use unique passwords for each system/service", "MANDATORY", "Password manager provided", "Yes", ""),
        ("Password Strength", "Create passwords meeting complexity requirements", "MANDATORY", "Technical enforcement", "Yes", ""),
        ("Secure Storage", "Never write down passwords or store in plain text", "MANDATORY", "Approved password manager", "Yes", ""),
        ("Compromise Reporting", "Immediately report suspected credential compromise", "MANDATORY", "Incident reporting process", "Yes", ""),
        ("MFA Device Security", "Protect MFA devices (phones, tokens) from unauthorised access", "MANDATORY", "Device management policy", "Yes", ""),
        ("Phishing Awareness", "Verify authenticity before entering credentials", "MANDATORY", "Security awareness training", "Yes", ""),
        ("Session Security", "Lock workstation when away, log out of shared systems", "MANDATORY", "Auto-lock policies", "Yes", ""),
        ("Personal Device MFA", "Secure personal devices used for MFA", "CONDITIONAL", "MDM for BYOD", "Yes", ""),
        ("Recovery Code Storage", "Store MFA recovery codes securely offline", "MANDATORY", "User guidance provided", "Yes", ""),
        ("Credential Sharing Prohibition", "Never use shared accounts for individual actions", "MANDATORY", "Audit logging", "Yes", ""),
        ("Employment Contract", "Acknowledge password responsibilities in contract", "MANDATORY", "HR onboarding process", "No", ""),
    ]

    row = 4
    for resp in responsibilities:
        for col, value in enumerate(resp, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col == 6:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"C4:C{row+5}", '"MANDATORY,RECOMMENDED,CONDITIONAL"')
    add_data_validation(ws, f"E4:E{row+5}", '"Yes,No,Completed"')

    set_column_widths(ws, {"A": 28, "B": 45, "C": 18, "D": 30, "E": 16, "F": 25})
    ws.freeze_panes = "A4"
    logger.info("Created User Responsibilities sheet")


def create_system_requirements_sheet(wb: Workbook):
    ws = wb.create_sheet("System Requirements")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "AUTHENTICATION SYSTEM TECHNICAL REQUIREMENTS"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"].value = f"{CONTROL_REF} | Technical requirements for authentication system implementation"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = CENTER_ALIGN

    headers = ["Requirement Area", "Technical Requirement", "Standard/Protocol", "Implementation", "Status", "Evidence", "Notes"]
    create_header_row(ws, 3, headers)

    requirements = [
        ("Password Hashing", "Use approved password hashing algorithms", "bcrypt (cost>=12), Argon2id, PBKDF2", "", "", "", ""),
        ("Salt Generation", "Unique cryptographic salt per password", "Min 128-bit random salt", "", "", "", ""),
        ("Transmission Security", "Encrypt credentials in transit", "TLS 1.2+ only", "", "", "", ""),
        ("Session Management", "Secure session token handling", "HttpOnly, Secure, SameSite cookies", "", "", "", ""),
        ("Account Lockout", "Implement progressive lockout", "5-10 attempts, 15+ min lockout", "", "", "", ""),
        ("Brute Force Protection", "Rate limiting on authentication", "Max 5 attempts/minute per IP", "", "", "", ""),
        ("Audit Logging", "Log all authentication events", "Success, failure, lockout, unlock", "", "", "", ""),
        ("Secure Recovery", "Secure password reset process", "Email verification, time-limited tokens", "", "", "", ""),
        ("MFA Integration", "Support for MFA protocols", "TOTP, WebAuthn/FIDO2, Push", "", "", "", ""),
        ("SSO Support", "Support for single sign-on", "SAML 2.0, OIDC, OAuth 2.0", "", "", "", ""),
        ("Password Exposure Check", "Check against breached password databases", "HIBP API integration", "", "", "", ""),
        ("Credential Isolation", "Separate credential stores by environment", "Prod/dev/test separation", "", "", "", ""),
        ("Secrets Management", "Centralised secrets management", "HashiCorp Vault, Azure Key Vault", "", "", "", ""),
        ("Certificate Management", "Automated certificate lifecycle", "ACME, internal PKI", "", "", "", ""),
    ]

    row = 4
    for req in requirements:
        for col, value in enumerate(req, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [4, 5, 6, 7]:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"E4:E{row+5}", '"Implemented,In Progress,Planned,Not Started,N/A"')

    set_column_widths(ws, {"A": 22, "B": 35, "C": 32, "D": 25, "E": 14, "F": 20, "G": 25})
    ws.freeze_panes = "A4"
    logger.info("Created System Requirements sheet")


def create_evidence_register(wb: Workbook):
    """Create GS-ER-compliant Evidence Register sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    # Row 1: Title banner
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = _border

    # Row 2: Italic subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = _border

    # Row 3: empty separator

    # Row 4: Column headers (003366 navy)
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border

    # Data validations
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other"',
        allow_blank=True,
    )
    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)
    ws.add_data_validation(ver_status_dv)

    # Row 5: Sample row (F2F2F2 grey)
    sample_data = {
        1: "EV-001", 2: "Assessment Area", 3: "Policy Document",
        4: "Example evidence description", 5: "\\\\fileserver\\policies\\document.pdf",
        6: "15.01.2026", 7: "Assessor Name", 8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = _border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center", wrap_text=True
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # Rows 6-105: Empty FFFFCC rows (100 rows)
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = _border
            cell.alignment = Alignment(
                horizontal="center" if col == 1 else "left",
                vertical="center", wrap_text=True
            )
            cell.value = None
        ev_type_dv.add(ws[f"C{row}"])
        ver_status_dv.add(ws[f"H{row}"])

    # Column widths and freeze panes
    for col, width in [("A", 15), ("B", 25), ("C", 22), ("D", 40), ("E", 45), ("F", 16), ("G", 20), ("H", 22)]:
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"
    logger.info("Created Evidence Register sheet")


def create_approval_sheet(wb: Workbook):
    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header (Row 1) — A1:E1 standard merge
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Control reference (Row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # DOCUMENT SUMMARY banner (Row 3)
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "DOCUMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Control Reference:", CONTROL_ID),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")"),
        ("Version:", "1.0"),
        ("Classification:", "INTERNAL"),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        ws.cell(row=row, column=1).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    row += 1  # gap

    # 3 Approver sections
    approvers = [
        ("PREPARED BY (POLICY OWNER)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"A{row}"].border = border
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border

    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")


def create_summary_dashboard_sheet(wb: Workbook):
    """Create Gold Standard Summary Dashboard sheet for A.5.17.1."""
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ffffcc_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # -------------------------------------------------------------------------
    # Row 1: Title banner — GS-SD-014: must contain "— SUMMARY DASHBOARD"
    # -------------------------------------------------------------------------
    ws.merge_cells("A1:G1")
    ws["A1"] = "AUTHENTICATION POLICY AND STANDARDS \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle — GS-SD italic, 003366, horizontal="left", NO wrap_text
    ws.merge_cells("A2:G2")
    ws["A2"] = f"ISO/IEC 27001:2022 \u2014 Control A.5.17: {CONTROL_NAME} | Policy and Standards Assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: empty spacer

    # -------------------------------------------------------------------------
    # TABLE 1: Assessment Area Compliance Overview
    # -------------------------------------------------------------------------
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = navy_fill
    for c in range(1, 8):
        ws.cell(row=4, column=c).border = border
    ws.row_dimensions[4].height = 20

    # TABLE 1 headers (row 5) — D9D9D9, 000000 bold
    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(t1_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # TABLE 1 data rows (rows 6-10)
    # area_configs: (name, total_formula, compliant_formula, partial_formula, noncompliant_formula, na_formula)
    area_configs = [
        (
            "Password Policy",
            "=COUNTA('Password Policy'!A4:A22)",
            "=COUNTIF('Password Policy'!E4:E22,\"Yes\")",
            "=COUNTIF('Password Policy'!E4:E22,\"Partial\")",
            "=COUNTIF('Password Policy'!E4:E22,\"No\")",
            "=COUNTIF('Password Policy'!E4:E22,\"N/A\")",
        ),
        (
            "MFA Requirements",
            "=COUNTA('MFA Requirements'!A4:A17)",
            "=COUNTIF('MFA Requirements'!E4:E17,\"Implemented\")",
            "=COUNTIF('MFA Requirements'!E4:E17,\"In Progress\")",
            "=COUNTIF('MFA Requirements'!E4:E17,\"Not Started\")",
            "=COUNTIF('MFA Requirements'!E4:E17,\"Exception\")",
        ),
        (
            "Credential Standards",
            "=COUNTA('Credential Standards'!A4:A14)",
            "0",
            "0",
            "0",
            "=COUNTA('Credential Standards'!A4:A14)",
        ),
        (
            "User Responsibilities",
            "=COUNTA('User Responsibilities'!A4:A15)",
            "=COUNTIF('User Responsibilities'!E4:E15,\"Completed\")",
            "=COUNTIF('User Responsibilities'!E4:E15,\"Yes\")",
            "=COUNTIF('User Responsibilities'!E4:E15,\"No\")",
            "0",
        ),
        (
            "System Requirements",
            "=COUNTA('System Requirements'!A4:A17)",
            "=COUNTIF('System Requirements'!E4:E17,\"Implemented\")",
            "=COUNTIF('System Requirements'!E4:E17,\"In Progress\")",
            "=COUNTIF('System Requirements'!E4:E17,\"Not Started\")",
            "=COUNTIF('System Requirements'!E4:E17,\"N/A\")",
        ),
    ]

    for i, (name, total_f, comp_f, part_f, nc_f, na_f) in enumerate(area_configs):
        row = 6 + i
        formulas = [name, total_f, comp_f, part_f, nc_f, na_f]
        for col, val in enumerate(formulas, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(color="000000")
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        # Compliance % (col G)
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        cell_g.number_format = "0.0%"
        cell_g.font = Font(color="000000")
        cell_g.border = border
        cell_g.alignment = Alignment(horizontal="center", vertical="center")

    # TABLE 1 TOTAL row (row 11)
    total_row = 11
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="000000")
    ws.cell(row=total_row, column=1).fill = grey_fill
    ws.cell(row=total_row, column=1).border = border
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}10)"
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    cell_g_total = ws.cell(row=total_row, column=7)
    cell_g_total.value = f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"
    cell_g_total.number_format = "0.0%"
    cell_g_total.font = Font(bold=True, color="000000")
    cell_g_total.fill = grey_fill
    cell_g_total.border = border
    cell_g_total.alignment = Alignment(horizontal="center", vertical="center")

    # -------------------------------------------------------------------------
    # TABLE 2: Key Performance Indicators (starts row 13)
    # -------------------------------------------------------------------------
    t2_banner_row = 13
    ws.merge_cells(f"A{t2_banner_row}:G{t2_banner_row}")
    ws[f"A{t2_banner_row}"] = "TABLE 2: KEY PERFORMANCE INDICATORS"
    ws[f"A{t2_banner_row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t2_banner_row}"].fill = navy_fill
    for c in range(1, 8):
        ws.cell(row=t2_banner_row, column=c).border = border
    ws.row_dimensions[t2_banner_row].height = 20

    # TABLE 2 headers (row 14) — D9D9D9, NOT 4472C4 (GS-SD-016)
    t2_hdr_row = 14
    t2_headers = ["Metric", "Value", "", "", "", "", ""]
    for col, hdr in enumerate(t2_headers, 1):
        cell = ws.cell(row=t2_hdr_row, column=col, value=hdr if hdr else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # TABLE 2 metrics (rows 15-29)
    metrics = [
        ("Password Policy: Requirements Assessed", "=COUNTA('Password Policy'!A4:A22)"),
        ("Password Policy: Requirements Compliant", "=COUNTIF('Password Policy'!E4:E22,\"Yes\")"),
        ("Password Policy: Non-Compliant Requirements", "=COUNTIF('Password Policy'!E4:E22,\"No\")"),
        ("Password Policy: Partially Met", "=COUNTIF('Password Policy'!E4:E22,\"Partial\")"),
        ("MFA: Mandatory Systems Defined", "=COUNTIF('MFA Requirements'!B4:B17,\"MANDATORY\")"),
        ("MFA: Recommended Systems Defined", "=COUNTIF('MFA Requirements'!B4:B17,\"RECOMMENDED\")"),
        ("MFA: Systems Fully Implemented", "=COUNTIF('MFA Requirements'!E4:E17,\"Implemented\")"),
        ("MFA: Systems Not Started", "=COUNTIF('MFA Requirements'!E4:E17,\"Not Started\")"),
        ("MFA: Exceptions Granted", "=COUNTIF('MFA Requirements'!E4:E17,\"Exception\")"),
        ("Credential Types Defined", "=COUNTA('Credential Standards'!A4:A14)"),
        ("User Responsibilities: Training Required", "=COUNTIF('User Responsibilities'!E4:E15,\"Yes\")"),
        ("User Responsibilities: Training Completed", "=COUNTIF('User Responsibilities'!E4:E15,\"Completed\")"),
        ("System Requirements: Implemented", "=COUNTIF('System Requirements'!E4:E17,\"Implemented\")"),
        ("System Requirements: In Progress", "=COUNTIF('System Requirements'!E4:E17,\"In Progress\")"),
        ("System Requirements: Not Started", "=COUNTIF('System Requirements'!E4:E17,\"Not Started\")"),
    ]

    row = 15
    for metric_label, metric_formula in metrics:
        cell_a = ws.cell(row=row, column=1, value=metric_label)
        cell_a.font = Font(color="000000")  # non-bold per GS-SD-015
        cell_a.border = border
        cell_a.alignment = Alignment(horizontal="left", vertical="center")
        cell_b = ws.cell(row=row, column=2, value=metric_formula)
        cell_b.font = Font(color="000000")
        cell_b.border = border
        cell_b.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = border
        row += 1

    # 2 empty buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
        row += 1

    # -------------------------------------------------------------------------
    # TABLE 3: Critical Findings (starts at row + 1)
    # -------------------------------------------------------------------------
    crit_start = row + 1
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{crit_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = red_fill
    for c in range(1, 8):
        ws.cell(row=crit_start, column=c).border = border
    ws.row_dimensions[crit_start].height = 20

    # TABLE 3 headers
    t3_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col, hdr in enumerate(t3_headers, 1):
        cell = ws.cell(row=crit_start + 1, column=col, value=hdr if hdr else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # TABLE 3 findings
    findings = [
        ("Password Policy", "Password requirements marked Non-Compliant",
         "=COUNTIF('Password Policy'!E4:E22,\"No\")", "Critical", "Immediate"),
        ("Password Policy", "Password requirements partially met",
         "=COUNTIF('Password Policy'!E4:E22,\"Partial\")", "High", "Urgent"),
        ("MFA Requirements", "MFA systems not yet started",
         "=COUNTIF('MFA Requirements'!E4:E17,\"Not Started\")", "Critical", "Immediate"),
        ("MFA Requirements", "MFA exceptions granted",
         "=COUNTIF('MFA Requirements'!E4:E17,\"Exception\")", "High", "Urgent"),
        ("System Requirements", "System requirements not started",
         "=COUNTIF('System Requirements'!E4:E17,\"Not Started\")", "High", "Urgent"),
        ("System Requirements", "System requirements in progress (incomplete)",
         "=COUNTIF('System Requirements'!E4:E17,\"In Progress\")", "Medium", "Plan"),
        ("User Responsibilities", "Training not completed",
         "=COUNTIF('User Responsibilities'!E4:E15,\"No\")", "Medium", "Plan"),
    ]

    row = crit_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(color="000000")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    # 2 empty buffer rows (FFFFCC)
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border
        row += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"
    logger.info("Created Summary Dashboard sheet")


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("=" * 70)
    logger.info(f"{DOCUMENT_ID} {WORKBOOK_NAME} Generator")
    logger.info("=" * 70)

    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.remove(wb.active)

    create_instructions_sheet(wb.create_sheet())
    create_password_policy_sheet(wb)
    create_mfa_requirements_sheet(wb)
    create_credential_standards_sheet(wb)
    create_user_responsibilities_sheet(wb)
    create_system_requirements_sheet(wb)
    create_evidence_register(wb)
    create_summary_dashboard_sheet(wb)
    create_approval_sheet(wb)


    finalize_validations(wb)
    wb.save(output_path)
    logger.info("=" * 70)
    logger.info(f"SUCCESS: Workbook saved as {OUTPUT_FILENAME}")
    logger.info("=" * 70)

    return 0


def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
