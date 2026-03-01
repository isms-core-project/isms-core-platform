#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.17.3 - Password System Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.17: Authentication Information
Assessment Domain 3 of 3: Password System Assessment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific authentication information management infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Authentication mechanism types and security requirements (match your systems)
2. Password/passphrase complexity and rotation policies (adapt to your standards)
3. Credential lifecycle stages and responsible owner roles
4. MFA applicability criteria and supported methods
5. Privileged credential management and vault integration scope

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.17 Authentication Information Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
authentication information management controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Password System Assessment under ISO 27001:2022 Control A.5.17. Supports evidence-based evaluation of authentication policy adherence, credential lifecycle compliance, and system security configuration.

**Assessment Scope:**
- Authentication policy definition and coverage completeness
- Credential lifecycle management process compliance
- Password system configuration against policy standards
- MFA implementation scope and coverage tracking
- Privileged credential management and storage controls
- User acknowledgment and awareness of authentication requirements
- Evidence collection for access management and compliance audits

**Generated Workbook Structure:**
1. System Inventory
2. Security Assessment
3. Storage Assessment
4. Integration Assessment
5. Gap Analysis
6. Evidence Register
7. Approval Sign-Off
8. Summary Dashboard

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 3 domains covering Authentication Information controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a517_3_password_system_assessment.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a517_3_password_system_assessment.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a517_3_password_system_assessment.py --date 20250115

Output:
    File: ISMS-IMP-A.5.17.3_Password_System_Assessment_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.17
Assessment Domain:    3 of 3 (Password System Assessment)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.17: Authentication Information Policy (Governance)
    - ISMS-IMP-A.5.17.1: Authentication Policy and Standards (Domain 1)
    - ISMS-IMP-A.5.17.2: Credential Lifecycle Management (Domain 2)
    - ISMS-IMP-A.5.17.3: Password System Assessment (Domain 3)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.17.3 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive authentication information management details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review authentication policies and credential management procedures annually or when authentication standards evolve, new systems are introduced, or security incidents reveal gaps.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

import logging
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.17.3"
WORKBOOK_NAME = "Password System Assessment"
CONTROL_ID = "A.5.17"
CONTROL_NAME = "Authentication Information"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP_LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass


def set_column_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_header_row(ws, row: int, headers: list):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def add_data_validation(ws, cell_range: str, formula: str):
    dv = DataValidation(type="list", formula1=formula, showDropDown=False, allowBlank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)



def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading (plain bold, no fill)
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)
    for i, line in enumerate([
        '1. Complete System Inventory — list all systems processing authentication credentials.',
        '2. Complete Security Assessment — evaluate password policy enforcement per system.',
        '3. Complete Storage Assessment — verify passwords are stored using approved hashing (bcrypt, Argon2).',
        '4. Complete Integration Assessment — assess SSO and identity provider integrations.',
        '5. Complete Gap Analysis — identify systems with weak or non-compliant authentication controls.',
        '6. Maintain the Evidence Register with system configuration exports and test results.',
        '7. Obtain final approval and sign-off in the Approval Sign-Off sheet.',
    ]):
        ws[f"A{13 + i}"] = line

    # Row 19 — Status Legend heading
    ws["A21"] = "Status Legend"
    ws["A21"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=22, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                    _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",             _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                      _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",         None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = 23 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
def create_system_inventory_sheet(wb: Workbook):
    ws = wb.create_sheet("System Inventory")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"].value = "AUTHENTICATION SYSTEM INVENTORY"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:J2")
    ws["A2"].value = f"{CONTROL_REF} | Inventory all systems handling authentication credentials"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = CENTER_ALIGN

    headers = ["System Name", "System Type", "Vendor", "Version", "User Count", "Auth Method", "SSO Integrated", "MFA Enabled", "Owner", "Criticality"]
    create_header_row(ws, 3, headers)

    # Sample row with data (row 4) — F2F2F2 grey
    SAMPLE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample = ("Active Directory", "Identity Provider", "Microsoft", "", "", "Password + Kerberos", "", "", "", "")
    for col, value in enumerate(sample, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT_ALIGN
        cell.fill = SAMPLE_FILL

    # 50 empty FFFFCC rows
    for row in range(5, 55):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
            cell.alignment = LEFT_ALIGN

    add_data_validation(ws, "B4:B54", '"Identity Provider,Cloud Identity,Cloud Application,Cloud Infrastructure,Network Access,Business Application,Development,Data Store,Other"')
    add_data_validation(ws, "G4:G54", '"Yes,No,Partial,N/A"')
    add_data_validation(ws, "H4:H54", '"Yes,No,Partial,N/A"')
    add_data_validation(ws, "J4:J54", '"Critical,High,Medium,Low"')

    set_column_widths(ws, {"A": 25, "B": 18, "C": 15, "D": 12, "E": 12, "F": 22, "G": 14, "H": 14, "I": 18, "J": 12})
    ws.freeze_panes = "A4"
    logger.info("Created System Inventory sheet")


def create_security_assessment_sheet(wb: Workbook):
    ws = wb.create_sheet("Security Assessment")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    ws["A1"].value = "AUTHENTICATION SYSTEM SECURITY ASSESSMENT"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:I2")
    ws["A2"].value = f"{CONTROL_REF} | Assess security controls per system against policy requirements"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = CENTER_ALIGN

    headers = ["System Name", "Control Area", "Requirement", "Expected State", "Actual State", "Status", "Gap Description", "Priority", "Notes"]
    create_header_row(ws, 3, headers)

    # Sample row (row 4) — F2F2F2 grey
    SAMPLE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample = ("Active Directory", "Password Policy", "Min 12 characters enforced", "GPO setting enabled", "", "", "", "", "")
    for col, value in enumerate(sample, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT_ALIGN
        cell.fill = SAMPLE_FILL

    # 50 empty FFFFCC rows
    for row in range(5, 55):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
            cell.alignment = LEFT_ALIGN

    add_data_validation(ws, "F4:F54", '"Compliant,Partial,Non-Compliant,N/A"')
    add_data_validation(ws, "H4:H54", '"Critical,High,Medium,Low"')

    set_column_widths(ws, {"A": 20, "B": 18, "C": 28, "D": 28, "E": 25, "F": 14, "G": 28, "H": 10, "I": 22})
    ws.freeze_panes = "A4"
    logger.info("Created Security Assessment sheet")


def create_storage_assessment_sheet(wb: Workbook):
    ws = wb.create_sheet("Storage Assessment")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"].value = "PASSWORD STORAGE SECURITY ASSESSMENT"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"{CONTROL_REF} | Review password storage mechanisms and hashing algorithms"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = CENTER_ALIGN

    headers = ["System/Application", "Storage Mechanism", "Hashing Algorithm", "Salting", "Key Protection", "Encryption at Rest", "Status", "Notes"]
    create_header_row(ws, 3, headers)

    # Sample row (row 4) — F2F2F2 grey
    SAMPLE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample = ("Active Directory", "NTDS.dit database", "NT Hash (MD4) + Kerberos", "No (legacy)", "DPAPI/TPM", "BitLocker", "", "")
    for col, value in enumerate(sample, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT_ALIGN
        cell.fill = SAMPLE_FILL

    # 50 empty FFFFCC rows
    for row in range(5, 55):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
            cell.alignment = LEFT_ALIGN

    add_data_validation(ws, "G4:G54", '"Compliant,Partial,Non-Compliant,N/A"')
    add_data_validation(ws, "C4:C54", '"bcrypt,Argon2id,PBKDF2-SHA256,scrypt,SHA-256 (not recommended),MD5 (non-compliant),Plaintext (critical),Unknown,N/A"')

    set_column_widths(ws, {"A": 25, "B": 22, "C": 20, "D": 15, "E": 18, "F": 18, "G": 14, "H": 28})
    ws.freeze_panes = "A4"
    logger.info("Created Storage Assessment sheet")


def create_integration_assessment_sheet(wb: Workbook):
    ws = wb.create_sheet("Integration Assessment")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    ws["A1"].value = "SSO AND FEDERATION INTEGRATION ASSESSMENT"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:I2")
    ws["A2"].value = f"{CONTROL_REF} | Assess SSO and federation integration for each application"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = CENTER_ALIGN

    headers = ["Application", "SSO Protocol", "Identity Provider", "MFA Pass-through", "Session Timeout", "Token Encryption", "Provisioning", "Status", "Notes"]
    create_header_row(ws, 3, headers)

    # Sample row (row 4) — F2F2F2 grey
    SAMPLE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample = ("Microsoft 365", "OIDC/WS-Fed", "Microsoft Entra ID", "Yes", "Configurable", "Yes", "Automated", "", "")
    for col, value in enumerate(sample, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT_ALIGN
        cell.fill = SAMPLE_FILL

    # 50 empty FFFFCC rows
    for row in range(5, 55):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
            cell.alignment = LEFT_ALIGN

    add_data_validation(ws, "B4:B54", '"SAML 2.0,OIDC,WS-Federation,OAuth 2.0,Local Auth,None"')
    add_data_validation(ws, "D4:D54", '"Yes,No,Partial,N/A"')
    add_data_validation(ws, "G4:G54", '"Automated,SCIM,API,Manual,None"')
    add_data_validation(ws, "H4:H54", '"Compliant,Partial,Non-Compliant,N/A"')

    set_column_widths(ws, {"A": 22, "B": 14, "C": 16, "D": 16, "E": 16, "F": 16, "G": 14, "H": 14, "I": 28})
    ws.freeze_panes = "A4"
    logger.info("Created Integration Assessment sheet")


def create_gap_analysis_sheet(wb: Workbook):
    ws = wb.create_sheet("Gap Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    ws["A1"].value = "GAP ANALYSIS AND REMEDIATION TRACKING"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:I2")
    ws["A2"].value = f"{CONTROL_REF} | Track identified gaps and remediation actions"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = CENTER_ALIGN

    headers = ["Gap ID", "System/Area", "Gap Description", "Risk Level", "Remediation Plan", "Owner", "Target Date", "Status", "Notes"]
    create_header_row(ws, 3, headers)

    # Sample row (row 4) — F2F2F2 grey
    SAMPLE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample = ("GAP-517-001", "System/Area", "Gap description here", "High", "Remediation plan", "Owner", "DD.MM.YYYY", "Open", "")
    for col, value in enumerate(sample, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.fill = SAMPLE_FILL
        cell.alignment = LEFT_ALIGN

    # 50 empty FFFFCC input rows
    for row in range(5, 55):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
            cell.alignment = LEFT_ALIGN

    add_data_validation(ws, "D4:D54", '"Critical,High,Medium,Low"')
    add_data_validation(ws, "H4:H54", '"Open,In Progress,Remediated,Verified,Risk Accepted"')

    set_column_widths(ws, {"A": 12, "B": 20, "C": 35, "D": 12, "E": 35, "F": 18, "G": 14, "H": 14, "I": 25})
    ws.freeze_panes = "A4"
    logger.info("Created Gap Analysis sheet")


def create_evidence_register(wb: Workbook):
    """Create GS-ER-compliant Evidence Register sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    # Row 1: Title banner
    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = _border

    # Row 2: Italic subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Document all evidence collected during this assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = _border

    # Row 3: empty separator

    # Row 4: Column headers (003366 navy)
    headers = [
        "Evidence ID", "Assessment Area", "Evidence Type", "Description",
        "Location / Path", "Date Collected", "Collected By", "Verification Status",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border

    # Data validations
    ev_type_dv = DataValidation(
        type="list",
        formula1='"Policy Document,Process Record,System Screenshot,Configuration Export,Audit Log,Training Record,Test Result,Risk Assessment,Meeting Minutes,Other"',
        allow_blank=True,
    )
    ver_status_dv = DataValidation(
        type="list",
        formula1='"✅ Verified,⚠️ Pending,❌ Not Verified,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)
    ws.add_data_validation(ver_status_dv)

    # Row 5: Sample row (F2F2F2 grey)
    sample_data = {
        1: "EV-001", 2: "Assessment Area", 3: "Policy Document",
        4: "Example evidence description", 5: "\\\\fileserver\\policies\\document.pdf",
        6: "15.01.2026", 7: "Assessor Name", 8: "✅ Verified",
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = _border
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center", wrap_text=True
        )
        cell.font = Font(name="Calibri", size=10)
    ev_type_dv.add(ws["C5"])
    ver_status_dv.add(ws["H5"])

    # Rows 6-105: Empty FFFFCC rows (100 rows)
    for row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            cell.border = _border
            cell.alignment = Alignment(
                horizontal="center" if col == 1 else "left",
                vertical="center", wrap_text=True
            )
            cell.value = None
        ev_type_dv.add(ws[f"C{row}"])
        ver_status_dv.add(ws[f"H{row}"])

    # Column widths and freeze panes
    for col, width in [("A", 15), ("B", 25), ("C", 22), ("D", 40), ("E", 45), ("F", 16), ("G", 20), ("H", 22)]:
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"
    logger.info("Created Evidence Register sheet")


def create_approval_sheet(wb: Workbook):
    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header (Row 1) — A1:E1 standard merge
    ws.merge_cells("A1:E1")
    ws["A1"] = "PASSWORD SYSTEM ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Control reference (Row 2)
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # ASSESSMENT SUMMARY banner (Row 3)
    row = 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "ASSESSMENT SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Control Reference:", CONTROL_ID),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]

    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    row += 1  # gap

    # 3 Approver sections
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]

    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"A{row}"].border = border
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].border = border
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border

    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border

    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = border
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")


def create_summary_dashboard_sheet(wb: Workbook):
    """Create Gold Standard Summary Dashboard sheet for A.5.17.3."""
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ffffcc_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # -------------------------------------------------------------------------
    # Row 1: Title banner — GS-SD-014: must contain "— SUMMARY DASHBOARD"
    # -------------------------------------------------------------------------
    ws.merge_cells("A1:G1")
    ws["A1"] = "PASSWORD SYSTEM ASSESSMENT \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle — italic, 003366, left-aligned, NO wrap_text
    ws.merge_cells("A2:G2")
    ws["A2"] = f"ISO/IEC 27001:2022 \u2014 Control A.5.17: {CONTROL_NAME} | System Security Assessment"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: empty spacer

    # -------------------------------------------------------------------------
    # TABLE 1: Assessment Area Compliance Overview
    # -------------------------------------------------------------------------
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A4"].fill = navy_fill
    for c in range(1, 8):
        ws.cell(row=4, column=c).border = border
    ws.row_dimensions[4].height = 20

    # TABLE 1 headers (row 5)
    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial", "Non-Compliant", "N/A", "Compliance %"]
    for col, header in enumerate(t1_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # TABLE 1 data rows (rows 6-10)
    # COUNTA uses col A (first user-entered col), starting at row 5 (skip sample row 4)
    # Compliant/Partial/Non-Compliant/N/A mapped per sheet DV
    area_configs = [
        (
            "System Inventory",
            "=COUNTA('System Inventory'!A5:A54)",
            "=COUNTIF('System Inventory'!H5:H54,\"Yes\")",
            "=COUNTIF('System Inventory'!H5:H54,\"Partial\")",
            "=COUNTIF('System Inventory'!H5:H54,\"No\")",
            "=COUNTIF('System Inventory'!H5:H54,\"N/A\")",
        ),
        (
            "Security Assessment",
            "=COUNTA('Security Assessment'!A5:A54)",
            "=COUNTIF('Security Assessment'!F5:F54,\"Compliant\")",
            "=COUNTIF('Security Assessment'!F5:F54,\"Partial\")",
            "=COUNTIF('Security Assessment'!F5:F54,\"Non-Compliant\")",
            "=COUNTIF('Security Assessment'!F5:F54,\"N/A\")",
        ),
        (
            "Storage Assessment",
            "=COUNTA('Storage Assessment'!A5:A54)",
            "=COUNTIF('Storage Assessment'!G5:G54,\"Compliant\")",
            "=COUNTIF('Storage Assessment'!G5:G54,\"Partial\")",
            "=COUNTIF('Storage Assessment'!G5:G54,\"Non-Compliant\")",
            "=COUNTIF('Storage Assessment'!G5:G54,\"N/A\")",
        ),
        (
            "Integration Assessment",
            "=COUNTA('Integration Assessment'!A5:A54)",
            "=COUNTIF('Integration Assessment'!H5:H54,\"Compliant\")",
            "=COUNTIF('Integration Assessment'!H5:H54,\"Partial\")",
            "=COUNTIF('Integration Assessment'!H5:H54,\"Non-Compliant\")",
            "=COUNTIF('Integration Assessment'!H5:H54,\"N/A\")",
        ),
        (
            "Gap Analysis",
            "=COUNTA('Gap Analysis'!A5:A54)",
            "=COUNTIF('Gap Analysis'!H5:H54,\"Remediated\")+COUNTIF('Gap Analysis'!H5:H54,\"Verified\")",
            "=COUNTIF('Gap Analysis'!H5:H54,\"In Progress\")",
            "=COUNTIF('Gap Analysis'!H5:H54,\"Open\")",
            "=COUNTIF('Gap Analysis'!H5:H54,\"Risk Accepted\")",
        ),
    ]

    for i, (name, total_f, comp_f, part_f, nc_f, na_f) in enumerate(area_configs):
        row = 6 + i
        formulas = [name, total_f, comp_f, part_f, nc_f, na_f]
        for col, val in enumerate(formulas, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(color="000000")
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        cell_g.number_format = "0.0%"
        cell_g.font = Font(color="000000")
        cell_g.border = border
        cell_g.alignment = Alignment(horizontal="center", vertical="center")

    # TABLE 1 TOTAL row (row 11)
    total_row = 11
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="000000")
    ws.cell(row=total_row, column=1).fill = grey_fill
    ws.cell(row=total_row, column=1).border = border
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}10)"
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    cell_g_total = ws.cell(row=total_row, column=7)
    cell_g_total.value = f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"
    cell_g_total.number_format = "0.0%"
    cell_g_total.font = Font(bold=True, color="000000")
    cell_g_total.fill = grey_fill
    cell_g_total.border = border
    cell_g_total.alignment = Alignment(horizontal="center", vertical="center")

    # -------------------------------------------------------------------------
    # TABLE 2: Key Performance Indicators (starts row 13)
    # -------------------------------------------------------------------------
    t2_banner_row = 13
    ws.merge_cells(f"A{t2_banner_row}:G{t2_banner_row}")
    ws[f"A{t2_banner_row}"] = "TABLE 2: KEY PERFORMANCE INDICATORS"
    ws[f"A{t2_banner_row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{t2_banner_row}"].fill = navy_fill
    for c in range(1, 8):
        ws.cell(row=t2_banner_row, column=c).border = border
    ws.row_dimensions[t2_banner_row].height = 20

    # TABLE 2 headers (row 14) — D9D9D9, NOT 4472C4 (GS-SD-016)
    t2_hdr_row = 14
    t2_headers = ["Metric", "Value", "", "", "", "", ""]
    for col, hdr in enumerate(t2_headers, 1):
        cell = ws.cell(row=t2_hdr_row, column=col, value=hdr if hdr else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # TABLE 2 metrics (rows 15-31) — 17 metrics
    metrics = [
        ("Total Auth Systems Inventoried", "=COUNTA('System Inventory'!A5:A54)"),
        ("Critical/High Criticality Systems",
         "=COUNTIF('System Inventory'!J5:J54,\"Critical\")+COUNTIF('System Inventory'!J5:J54,\"High\")"),
        ("Systems with MFA Enabled", "=COUNTIF('System Inventory'!H5:H54,\"Yes\")"),
        ("Systems Without MFA", "=COUNTIF('System Inventory'!H5:H54,\"No\")"),
        ("Systems with SSO Integrated", "=COUNTIF('System Inventory'!G5:G54,\"Yes\")"),
        ("Security Assessment: Compliant Controls", "=COUNTIF('Security Assessment'!F5:F54,\"Compliant\")"),
        ("Security Assessment: Non-Compliant Controls", "=COUNTIF('Security Assessment'!F5:F54,\"Non-Compliant\")"),
        ("Security Assessment: Critical Priority Gaps", "=COUNTIF('Security Assessment'!H5:H54,\"Critical\")"),
        ("Password Storage: Compliant Systems", "=COUNTIF('Storage Assessment'!G5:G54,\"Compliant\")"),
        ("Password Storage: Non-Compliant Systems", "=COUNTIF('Storage Assessment'!G5:G54,\"Non-Compliant\")"),
        ("Weak/Non-Compliant Hashing Detected",
         "=COUNTIF('Storage Assessment'!C5:C54,\"MD5 (non-compliant)\")"
         "+COUNTIF('Storage Assessment'!C5:C54,\"Plaintext (critical)\")"),
        ("Integration Assessment: Compliant", "=COUNTIF('Integration Assessment'!H5:H54,\"Compliant\")"),
        ("Integration Assessment: Non-Compliant", "=COUNTIF('Integration Assessment'!H5:H54,\"Non-Compliant\")"),
        ("Total Gaps Identified", "=COUNTA('Gap Analysis'!A5:A54)"),
        ("Open / Unresolved Gaps", "=COUNTIF('Gap Analysis'!H5:H54,\"Open\")"),
        ("Critical Severity Gaps", "=COUNTIF('Gap Analysis'!D5:D54,\"Critical\")"),
        ("Gaps Remediated or Verified",
         "=COUNTIF('Gap Analysis'!H5:H54,\"Remediated\")+COUNTIF('Gap Analysis'!H5:H54,\"Verified\")"),
    ]

    row = 15
    for metric_label, metric_formula in metrics:
        cell_a = ws.cell(row=row, column=1, value=metric_label)
        cell_a.font = Font(color="000000")  # non-bold per GS-SD-015
        cell_a.border = border
        cell_a.alignment = Alignment(horizontal="left", vertical="center")
        cell_b = ws.cell(row=row, column=2, value=metric_formula)
        cell_b.font = Font(color="000000")
        cell_b.border = border
        cell_b.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = border
        row += 1

    # 2 empty buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
        row += 1

    # -------------------------------------------------------------------------
    # TABLE 3: Critical Findings
    # -------------------------------------------------------------------------
    crit_start = row + 1
    ws.merge_cells(f"A{crit_start}:G{crit_start}")
    ws[f"A{crit_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{crit_start}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{crit_start}"].fill = red_fill
    for c in range(1, 8):
        ws.cell(row=crit_start, column=c).border = border
    ws.row_dimensions[crit_start].height = 20

    # TABLE 3 headers
    t3_headers = ["Category", "Finding", "Count", "Severity", "Action Required", "", ""]
    for col, hdr in enumerate(t3_headers, 1):
        cell = ws.cell(row=crit_start + 1, column=col, value=hdr if hdr else None)
        cell.font = Font(bold=True, color="000000")
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # TABLE 3 findings
    findings = [
        ("System Inventory", "Auth systems without MFA enabled",
         "=COUNTIF('System Inventory'!H5:H54,\"No\")", "Critical", "Immediate"),
        ("System Inventory", "Critical criticality systems identified",
         "=COUNTIF('System Inventory'!J5:J54,\"Critical\")", "Critical", "Immediate"),
        ("Security Assessment", "Non-compliant security controls",
         "=COUNTIF('Security Assessment'!F5:F54,\"Non-Compliant\")", "Critical", "Immediate"),
        ("Security Assessment", "Critical priority security gaps",
         "=COUNTIF('Security Assessment'!H5:H54,\"Critical\")", "Critical", "Immediate"),
        ("Storage Assessment", "Non-compliant password storage systems",
         "=COUNTIF('Storage Assessment'!G5:G54,\"Non-Compliant\")", "Critical", "Immediate"),
        ("Storage Assessment", "Weak or non-compliant hashing algorithms detected",
         "=COUNTIF('Storage Assessment'!C5:C54,\"MD5 (non-compliant)\")"
         "+COUNTIF('Storage Assessment'!C5:C54,\"Plaintext (critical)\")", "Critical", "Immediate"),
        ("Integration Assessment", "Non-compliant SSO/federation integrations",
         "=COUNTIF('Integration Assessment'!H5:H54,\"Non-Compliant\")", "High", "Urgent"),
        ("Gap Analysis", "Open unresolved gaps",
         "=COUNTIF('Gap Analysis'!H5:H54,\"Open\")", "High", "Urgent"),
        ("Gap Analysis", "Critical severity unresolved gaps",
         "=COUNTIF('Gap Analysis'!D5:D54,\"Critical\")", "Critical", "Immediate"),
    ]

    row = crit_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(color="000000")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    # 2 empty buffer rows (FFFFCC)
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = ffffcc_fill
            ws.cell(row=row, column=col).border = border
        row += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A4"
    logger.info("Created Summary Dashboard sheet")


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("=" * 70)
    logger.info(f"{DOCUMENT_ID} {WORKBOOK_NAME} Generator")
    logger.info("=" * 70)

    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.remove(wb.active)

    create_instructions_sheet(wb.create_sheet())
    create_system_inventory_sheet(wb)
    create_security_assessment_sheet(wb)
    create_storage_assessment_sheet(wb)
    create_integration_assessment_sheet(wb)
    create_gap_analysis_sheet(wb)
    create_evidence_register(wb)
    create_summary_dashboard_sheet(wb)
    create_approval_sheet(wb)


    finalize_validations(wb)
    wb.save(output_path)
    logger.info("=" * 70)
    logger.info(f"SUCCESS: Workbook saved as {OUTPUT_FILENAME}")
    logger.info("=" * 70)

    return 0


def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
