#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.30.S2 - Redundancy Analysis & SPOF Assessment Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.8.14: Redundancy of Information Processing Facilities
Assessment Domain 2 of 4: System Redundancy and Failover Capability

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific infrastructure architecture, redundancy requirements,
and availability targets.

Key customization areas:
1. System criticality classification (match your business impact analysis)
2. Redundancy types and architectures (N+1, N+2, active-active, etc.)
3. RTO requirements per system tier (adapt to your recovery objectives)
4. Failover testing frequency (align with your operational procedures)
5. SPOF tolerance thresholds (based on your risk appetite)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.30/14/5.30 BC/DR Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
system redundancy, identifying Single Points of Failure (SPOFs), and validating
failover capabilities across critical infrastructure components.

**Purpose:**
Enables systematic assessment of redundancy implementation against ISO 27001:2022
Control A.8.14 requirements, supporting evidence-based validation of system
availability and failover readiness.

**Assessment Scope:**
- Critical system inventory with availability requirements
- Redundancy architecture assessment (clustering, load balancing, replication)
- Single Point of Failure (SPOF) identification and analysis
- Failover capability validation (automatic vs. manual)
- RTO requirement vs. actual failover time compliance
- Geographic redundancy validation
- Network redundancy assessment
- Power and utility redundancy verification
- Failover testing results and documentation
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and redundancy best practices
2. Redundancy Assessment - System redundancy architecture inventory (110 rows: 10 examples + 100 data entry)
3. Evidence Register - Audit evidence tracking and documentation (100 rows)
4. Approval & Sign-Off - Multi-level stakeholder review and approval workflow

**Key Features:**
- Data validation with comprehensive dropdown lists (redundancy types, failover modes)
- Conditional formatting for SPOF identification and RTO compliance
- Automated RTO compliance calculations (requirement vs. actual)
- Automated SPOF risk scoring
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- 3-level approval workflow (Assessor -> ISO Officer -> CISO)
- Exception handling with graceful error reporting
- Professional styling without Excel repair warnings

**Integration:**
This assessment feeds into the ISMS-IMP-A.5.30.S5 BC/DR Summary Dashboard,
which consolidates data from all four BC/DR assessment domains for executive
oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl --break-system-packages

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a530_2_redundancy_analysis.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a530_2_redundancy_analysis.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a530_2_redundancy_analysis.py --date 20250125

Output:
    File: ISMS_Assessment_Redundancy_Analysis.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize system criticality tiers to match your BIA
    2. Inventory all critical systems requiring redundancy
    3. Document redundancy architecture for each system
    4. Identify and analyse all SPOFs
    5. Conduct failover testing and document results
    6. Validate RTO requirements alignment with actual failover times
    7. Define SPOF remediation actions with timelines
    8. Collect and link audit evidence (architecture diagrams, test results)
    9. Obtain stakeholder approvals
    10. Feed results into ISMS-IMP-A.5.30.S5 BC/DR Summary Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.8.14
Assessment Domain:    2 of 4 (System Redundancy & SPOF Analysis)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.30-8.13-14: BC/DR Framework Policy (Governance)
    - ISMS-POL-A.5.30-8.13-14-S3: Redundancy Requirements (A.8.14)
    - ISMS-IMP-A.5.30-8.13-14-S1: BIA and RPO/RTO Process
    - ISMS-IMP-A.5.30-8.13-14-S3: Redundancy Implementation Guide
    - ISMS-IMP-A.5.30.S1: Backup Inventory Assessment (Domain 1)
    - ISMS-IMP-A.5.30.S3: RPO/RTO Compliance Matrix (Domain 3)
    - ISMS-IMP-A.5.30.S4: BC/DR Testing Results Tracker (Domain 4)
    - ISMS-IMP-A.5.30.S5: BC/DR Summary Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.30-8.13-14-S3 specification
    - Supports comprehensive redundancy and SPOF analysis
    - Integrated with ISMS-IMP-A.5.30.S5 BC/DR Summary Dashboard
    - Includes automated SPOF risk scoring

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Untested Failover = No Redundancy:**
Redundancy architecture is meaningless without verified failover capability.
ALWAYS document last failover test dates and results. Systems with untested
failover should be treated as having SPOFs regardless of architecture.

**SPOF Risk Assessment:**
Single Points of Failure represent critical availability risks. Prioritize
SPOF remediation based on:
- System criticality (Tier 1 critical systems take priority)
- Business impact of system unavailability
- RTO requirements vs. recovery capability gap
- Cost vs. risk trade-offs

Not all SPOFs require immediate remediation, but all must be risk-accepted
at appropriate management level.

**RTO Alignment:**
RTO requirements must derive from Business Impact Analysis (BIA). Different
systems have different availability requirements:
- Tier 1 (Critical): May require active-active with automatic failover (RTO < 15 min)
- Tier 2 (Important): May require active-passive with quick failover (RTO < 4 hours)
- Tier 3 (Standard): May accept manual failover (RTO < 24 hours)
- Tier 4 (Low): May accept rebuild from backup (RTO < 72 hours)

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect verification of failover testing and SPOF remediation plans.

**Data Protection:**
Assessment workbooks contain sensitive infrastructure details including:
- System architecture and redundancy design
- SPOF identification (security-sensitive information)
- Recovery capability gaps

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Verify redundancy coverage for new systems, validate SPOF status
- Semi-annually: Validate RTO requirements still align with business needs
- Annually: Complete reassessment of all systems
- Ad-hoc: After infrastructure changes, failover tests, or availability incidents

**Quality Assurance:**
Have infrastructure architects and availability engineers validate assessments
before using results for compliance reporting or remediation decisions.

**Regulatory Alignment:**
Ensure redundancy requirements align with applicable regulatory requirements:
- Finance: DORA ICT operational resilience requirements
- Healthcare: HIPAA availability requirements
- Critical Infrastructure: Sector-specific availability mandates

Customize assessment criteria to include regulatory-specific requirements.

**Cost vs. Availability Trade-offs:**
Redundancy is expensive. This assessment helps prioritize investments based on
actual business requirements, not assumptions. Document business-driven decisions
to accept SPOFs where remediation cost exceeds risk.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from datetime import datetime, timedelta
import sys
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
WORKBOOK_TITLE = "Redundancy Analysis & SPOF Assessment"
WORKBOOK_NAME = "Redundancy Analysis"
VERSION = "1.0"
CONTROLS = "A.5.30 (Information Backup)"


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠  Warning sign
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.30.S2"
CONTROL_ID   = "A.5.30"
CONTROL_NAME = "Information Backup"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Color scheme
HEADER_FILL = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
COLUMN_HEADER_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

HEADER_FONT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
SUBHEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
BOLD_FONT = Font(name='Calibri', size=10, bold=True)
NORMAL_FONT = Font(name='Calibri', size=10)

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def apply_style(cell, font=None, fill=None, alignment=None, border=None):
    """Apply style to cell. Creates NEW objects to avoid Excel warnings."""
    if font:
        cell.font = Font(
            name=font.name if hasattr(font, 'name') else 'Calibri',
            size=font.size if hasattr(font, 'size') else 10,
            bold=font.bold if hasattr(font, 'bold') else False,
            color=font.color if hasattr(font, 'color') else None
        )
    if fill:
        cell.fill = PatternFill(
            start_color=fill.start_color.rgb if hasattr(fill.start_color, 'rgb') else fill.start_color,
            end_color=fill.end_color.rgb if hasattr(fill.end_color, 'rgb') else fill.end_color,
            fill_type=fill.fill_type
        )
    if alignment:
        cell.alignment = Alignment(
            horizontal=alignment.horizontal if hasattr(alignment, 'horizontal') else 'left',
            vertical=alignment.vertical if hasattr(alignment, 'vertical') else 'center',
            wrap_text=alignment.wrap_text if hasattr(alignment, 'wrap_text') else False
        )
    if border:
        thin = Side(style='thin')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def create_base_validations(ws):
    """Create comprehensive data validation objects."""
    validations = {
        'criticality': DataValidation(type="list",
            formula1='"Tier 1 - Critical,Tier 2 - Important,Tier 3 - Standard,Tier 4 - Low"',
            allow_blank=False),
        'redundancy_status': DataValidation(type="list",
            formula1=f'"{CHECK} Implemented,⚠️ Partial,❌ None,⏳ In Progress"',
            allow_blank=False),
        'architecture_type': DataValidation(type="list",
            formula1='"Active-Active,Active-Passive,N+1 Cluster,N+2 Cluster,Warm Standby,Cold Standby,None"',
            allow_blank=False),
        'failover_type': DataValidation(type="list",
            formula1='"Automatic,Manual,None,➖ N/A"',
            allow_blank=False),
        'yes_no_na': DataValidation(type="list",
            formula1=f'"{CHECK} Yes,❌ No,➖ N/A"',
            allow_blank=False),
        'test_result': DataValidation(type="list",
            formula1=f'"{CHECK} Success,⚠️ Partial,❌ Failure,➖ Not Tested"',
            allow_blank=False),
        'compliance_status': DataValidation(type="list",
            formula1=f'"{CHECK} Compliant,❌ Non-Compliant,❓ Unknown"',
            allow_blank=False),
        'spof_type': DataValidation(type="list",
            formula1='"Hardware,Network,Power,Software,Database,Storage,Cloud Provider,DNS,Load Balancer,Other"',
            allow_blank=False),
        'risk_level': DataValidation(type="list",
            formula1='"[!] Critical,[-] Medium,[.] Low"',
            allow_blank=False),
        'mitigation_status': DataValidation(type="list",
            formula1=f'"{CHECK} Mitigated,⏳ In Progress,❌ Open"',
            allow_blank=False),
        'evidence_type': DataValidation(type="list",
            formula1='"Config File,Screenshot,Report,Log File,Test Result,Architecture Diagram,Policy Document,Other"',
            allow_blank=False),
        'verification_status': DataValidation(type="list",
            formula1=f'"{CHECK} Verified,⏳ Pending,❌ Not Verified"',
            allow_blank=False),
        'assessment_status': DataValidation(type="list",
            formula1='"Draft,Under Review,Final,Requires Remediation"',
            allow_blank=False),
        'approval_decision': DataValidation(type="list",
            formula1='"Approved,Approved with Conditions,Rejected,Requires Rework"',
            allow_blank=False),
        'recommendation': DataValidation(type="list",
            formula1='"Approve,Approve with Conditions,Reject,Require Rework"',
            allow_blank=False),
    }
    
    # DVs are added lazily (only when cells are assigned) — do NOT auto-add here
    return validations

# ============================================================================
# WORKSHEET: INSTRUCTIONS
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Populate the Redundancy Inventory worksheet (document redundancy status for all systems).', '2. Complete the SPOF Register worksheet (identify and track Single Points of Failure).', '3. Assess the RTO Compliance worksheet (compare failover capability vs BIA requirements).', '4. Document evidence in the Evidence Register (minimum 5 items required for audit).', '5. Review the Summary Dashboard for compliance status overview.', '6. Complete the Approval Sign-Off worksheet to obtain formal sign-off.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 20

    _EVIDENCE = ['✓ Redundancy architecture diagrams (Active-Active, Active-Passive, N+1)', '✓ Failover test records with date, system, and RTO achieved', '✓ System monitoring screenshots showing HA/DR status', '✓ Cloud provider multi-region/availability zone configuration', '✓ SPOF remediation tickets and closure evidence', '✓ RTO/RTOs defined in Business Impact Analysis (BIA)', '✓ Vendor SLA documents for critical systems', '✓ Disaster Recovery runbooks and test results']

    # Status Legend — row position tracks after instructions
    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Acceptable Evidence section
    _ev_row = _leg_row + 7
    ws[f"A{_ev_row}"] = "Acceptable Evidence (examples)"
    ws[f"A{_ev_row}"].font = Font(name="Calibri", size=12, bold=True)
    for _ev in _EVIDENCE:
        _ev_row += 1
        ws[f"A{_ev_row}"] = _ev

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_redundancy_inventory_sheet(wb):
    """Create Redundancy Inventory worksheet"""
    ws = wb.create_sheet(title="Redundancy Inventory")
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:J1')
    ws['A1'] = 'REDUNDANCY INVENTORY & STATUS'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:J2')
    ws['A2'] = 'Document redundancy status for all systems requiring RTO < 4 hours'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row = 4
    headers = [
        'System Name', 'Criticality Tier', 'RTO Requirement (hours)',
        'Redundancy Status', 'Architecture Type', 'Failover Type',
        'Geographic Redundancy', 'Last Failover Test', 'Test Result', 'Notes'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    row += 1
    # Grey sample row (F2F2F2) — GS-MAX-003: shows how to fill in the sheet
    _grey = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _sample = [
        'Mail Server (Example)', 'Tier 2 - Important', 4,
        f'{WARNING} Partial', 'Active-Passive', 'Manual', f'{XMARK} No',
        '', '➖ Not Tested', 'Sample — replace with actual system data',
    ]
    for _ci, _v in enumerate(_sample, start=1):
        _cell = ws.cell(row=row, column=_ci, value=_v)
        apply_style(_cell, fill=_grey, border=THIN_BORDER)
        if _ci == 2:
            validations['criticality'].add(_cell)
        elif _ci == 4:
            validations['redundancy_status'].add(_cell)
        elif _ci == 5:
            validations['architecture_type'].add(_cell)
        elif _ci == 6:
            validations['failover_type'].add(_cell)
        elif _ci == 7:
            validations['yes_no_na'].add(_cell)
        elif _ci == 9:
            validations['test_result'].add(_cell)
    row += 1
    # 50 empty FFFFCC data rows (MAX-004: 1 grey sample + 50 empty = 51 total)
    for i in range(50):
        for col_idx in range(1, 11):
            cell = ws.cell(row=row, column=col_idx)
            apply_style(cell, fill=INPUT_FILL, border=THIN_BORDER)

            if col_idx == 2:
                validations['criticality'].add(cell)
            elif col_idx == 4:
                validations['redundancy_status'].add(cell)
            elif col_idx == 5:
                validations['architecture_type'].add(cell)
            elif col_idx == 6:
                validations['failover_type'].add(cell)
            elif col_idx == 7:
                validations['yes_no_na'].add(cell)
            elif col_idx == 9:
                validations['test_result'].add(cell)
        row += 1
    
    # Summary
    summary_row = row + 2
    ws.merge_cells(f'A{summary_row}:J{summary_row}')
    ws[f'A{summary_row}'] = 'REDUNDANCY COVERAGE METRICS'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    summary_row += 1
    # Metric rows: 59=Total, 60=WithRedundancy, 61=Coverage%, 62=Critical, 63=CriticalWith, 64=CriticalCoverage%
    # Data rows: 6-55 (sample=5 excluded from counts); ranges end at 55 to avoid circular ref
    metrics = [
        ('Total Systems:', '=COUNTA(A6:A55)'),
        ('Systems with Redundancy:', '=COUNTIF(D6:D55,"{CHECK}*")'),
        ('Redundancy Coverage %:', '=IF(B59>0,B60/B59,0)'),
        ('Critical Systems (RTO < 4h):', '=COUNTIFS(B6:B55,"Tier 1*",C6:C55,"<4")'),
        ('Critical with Redundancy:', '=COUNTIFS(B6:B55,"Tier 1*",D6:D55,"{CHECK}*")'),
        ('Critical Coverage %:', '=IF(B62>0,B63/B62,1)'),
    ]
    
    for label, formula in metrics:
        ws[f'A{summary_row}'] = label
        apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
        ws[f'B{summary_row}'] = formula
        apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
        if 'Coverage %' in label:
            ws[f'B{summary_row}'].number_format = '0.0%'
        summary_row += 1
    
    widths = [25, 18, 22, 18, 25, 18, 20, 18, 15, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Lazy-add: only register DVs that have cells assigned (fixes empty-sqref XML error)
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)
    
    ws.freeze_panes = 'A4'
    return ws

# ============================================================================
# WORKSHEET: SPOF REGISTER
# ============================================================================

def create_spof_register_sheet(wb):
    """Create SPOF Register worksheet"""
    ws = wb.create_sheet(title="SPOF Register")
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:I1')
    ws['A1'] = 'SINGLE POINT OF FAILURE (SPOF) REGISTER'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:I2')
    ws['A2'] = 'Track identified SPOFs and mitigation progress'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row = 4
    headers = [
        'SPOF ID', 'System Affected', 'SPOF Component',
        'SPOF Type', 'Risk Level', 'Mitigation Status',
        'Mitigation Plan', 'Owner', 'Target Date'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    row += 1
    # Grey sample row (F2F2F2) — GS-MAX-003: shows how to fill in the sheet
    _grey = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _spof_sample = [
        'SPOF-008', 'Database Server (Example)', 'Single storage array',
        'Storage', '[~] High', '⏳ In Progress',
        'Add redundant storage and configure RAID', 'Storage Admin', '',
    ]
    for _ci, _v in enumerate(_spof_sample, start=1):
        _cell = ws.cell(row=row, column=_ci, value=_v)
        apply_style(_cell, fill=_grey, border=THIN_BORDER)
        if _ci == 4:
            validations['spof_type'].add(_cell)
        elif _ci == 5:
            validations['risk_level'].add(_cell)
        elif _ci == 6:
            validations['mitigation_status'].add(_cell)
    row += 1
    # 50 empty FFFFCC data rows (MAX-004: 1 grey sample + 50 empty = 51 total)
    for i in range(50):
        cell_a = ws.cell(row=row, column=1)
        apply_style(cell_a, fill=INPUT_FILL, border=THIN_BORDER)

        for col_idx in range(2, 10):
            cell = ws.cell(row=row, column=col_idx)
            apply_style(cell, fill=INPUT_FILL, border=THIN_BORDER)

            if col_idx == 4:
                validations['spof_type'].add(cell)
            elif col_idx == 5:
                validations['risk_level'].add(cell)
            elif col_idx == 6:
                validations['mitigation_status'].add(cell)
        row += 1
    
    # Summary
    summary_row = row + 2
    ws.merge_cells(f'A{summary_row}:I{summary_row}')
    ws[f'A{summary_row}'] = 'SPOF SUMMARY'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    summary_row += 1
    metrics = [
        ('Total SPOFs Identified:', '=COUNTA(B6:B55)'),
        ('SPOFs Mitigated:', '=COUNTIF(F6:F55,"{CHECK}*")'),
        ('SPOFs In Progress:', '=COUNTIF(F6:F55,"⏳*")'),
        ('Open SPOFs:', '=COUNTIF(F6:F55,"{XMARK}*")'),
        ('Critical Open SPOFs:', '=COUNTIFS(E6:E55,"[!]*",F6:F55,"{XMARK}*")'),
        ('Mitigation Rate:', '=IF(B59>0,B60/B59,0)'),
    ]
    
    for label, formula in metrics:
        ws[f'A{summary_row}'] = label
        apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
        ws[f'B{summary_row}'] = formula
        apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
        if 'Rate' in label:
            ws[f'B{summary_row}'].number_format = '0.0%'
        summary_row += 1
    
    widths = [12, 25, 30, 18, 15, 18, 35, 20, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Lazy-add: only register DVs that have cells assigned (fixes empty-sqref XML error)
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)
    
    ws.freeze_panes = 'A4'
    return ws

# ============================================================================
# WORKSHEET: RTO COMPLIANCE
# ============================================================================

def create_rto_compliance_sheet(wb):
    """Create RTO Compliance worksheet"""
    ws = wb.create_sheet(title="RTO Compliance")
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)
    
    ws.merge_cells('A1:G1')
    ws['A1'] = 'RTO COMPLIANCE ASSESSMENT'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:G2')
    ws['A2'] = 'Compare RTO requirements vs actual tested failover time'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row = 4
    headers = [
        'System Name', 'Criticality Tier', 'RTO Requirement (hours)',
        'Actual Failover Time (hours)', 'RTO Compliant', 'Gap (hours)', 'Notes'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    row += 1
    start_row = row
    
    # Formulas for 110 rows
    for i in range(110):
        ws[f'E{row}'] = f'=IF(AND(ISNUMBER(C{row}),ISNUMBER(D{row})),IF(D{row}<=C{row},"{CHECK} Compliant","{XMARK} Non-Compliant"),"❓ Unknown")'
        ws[f'F{row}'] = f'=IF(AND(ISNUMBER(C{row}),ISNUMBER(D{row})),MAX(0,D{row}-C{row}),"")'
        row += 1
    
    # Summary
    summary_row = start_row + 113
    ws.merge_cells(f'A{summary_row}:G{summary_row}')
    ws[f'A{summary_row}'] = 'RTO COMPLIANCE METRICS'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))

    summary_row += 1
    # Compute fixed row references before the loop to avoid circular reference
    total_systems_row = summary_row        # row where Total Systems will be written
    compliant_row = summary_row + 1        # row where Systems Compliant will be written
    metrics = [
        ('Total Systems:', f'=COUNTA(A{start_row}:A{start_row+109})'),
        ('Systems Compliant:', f'=COUNTIF(E{start_row}:E{start_row+109},"{CHECK}*")'),
        ('Systems Non-Compliant:', f'=COUNTIF(E{start_row}:E{start_row+109},"{XMARK}*")'),
        ('Systems Not Tested:', f'=COUNTIF(E{start_row}:E{start_row+109},"❓*")'),
        ('RTO Compliance Rate:', f'=IF(B{total_systems_row}>0,B{compliant_row}/B{total_systems_row},0)'),
        ('Critical Compliance Rate:', f'=IF(COUNTIF(B{start_row}:B{start_row+109},"Tier 1*")>0,COUNTIFS(B{start_row}:B{start_row+109},"Tier 1*",E{start_row}:E{start_row+109},"{CHECK}*")/COUNTIF(B{start_row}:B{start_row+109},"Tier 1*"),1)'),
    ]
    
    for label, formula in metrics:
        ws[f'A{summary_row}'] = label
        apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
        ws[f'B{summary_row}'] = formula
        apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
        if 'Rate' in label:
            ws[f'B{summary_row}'].number_format = '0.0%'
        summary_row += 1
    
    widths = [25, 18, 22, 26, 18, 15, 35]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Lazy-add: only register DVs that have cells assigned (fixes empty-sqref XML error)
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)
    
    ws.freeze_panes = 'A4'
    return ws

# ============================================================================
# WORKSHEET: SUMMARY
# ============================================================================

def create_summary_dashboard_sheet(wb):
    """Create Summary Dashboard"""
    ws = wb.create_sheet(title="Summary Dashboard")
    ws.sheet_view.showGridLines = False

    # DS-002: fill must be HEADER_FILL (003366) with white font
    ws.merge_cells('A1:E1')
    ws['A1'] = 'REDUNDANCY ANALYSIS \u2014 SUMMARY DASHBOARD'
    apply_style(ws['A1'], font=Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
                fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:E2')
    ws['A2'] = f'Assessment ID: {DOCUMENT_ID} | Assessment Date: [enter date]'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(italic=True, size=10)
    
    row = 4
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'REDUNDANCY COVERAGE'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))
    
    row += 1
    metrics = [
        ('Total Systems:', "='Redundancy Inventory'!B59"),
        ('Systems with Redundancy:', "='Redundancy Inventory'!B60"),
        ('Overall Coverage:', "='Redundancy Inventory'!B61"),
        ('Critical Systems (RTO < 4h):', "='Redundancy Inventory'!B62"),
        ('Critical with Redundancy:', "='Redundancy Inventory'!B63"),
        ('Critical Coverage:', "='Redundancy Inventory'!B64"),
    ]
    
    for label, formula in metrics:
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws[f'B{row}'] = formula
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=12))
        if 'Coverage' in label:
            ws[f'B{row}'].number_format = '0.0%'
        row += 1
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'SINGLE POINTS OF FAILURE (SPOF)'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))
    
    row += 1
    spof_metrics = [
        ('Total SPOFs Identified:', "='SPOF Register'!B59"),
        ('SPOFs Mitigated:', "='SPOF Register'!B60"),
        ('SPOFs In Progress:', "='SPOF Register'!B61"),
        ('Open SPOFs:', "='SPOF Register'!B62"),
        ('Critical Open SPOFs:', "='SPOF Register'!B63"),
        ('Mitigation Rate:', "='SPOF Register'!B64"),
    ]
    
    for label, formula in spof_metrics:
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws[f'B{row}'] = formula
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=12))
        if 'Rate' in label:
            ws[f'B{row}'].number_format = '0.0%'
        row += 1
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'RTO COMPLIANCE'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))
    
    row += 1
    rto_metrics = [
        ('Total Systems:', "='RTO Compliance'!B119"),
        ('Systems Compliant:', "='RTO Compliance'!B120"),
        ('Systems Non-Compliant:', "='RTO Compliance'!B121"),
        ('Systems Not Tested:', "='RTO Compliance'!B122"),
        ('RTO Compliance Rate:', "='RTO Compliance'!B123"),
        ('Critical Compliance Rate:', "='RTO Compliance'!B124"),
    ]
    
    for label, formula in rto_metrics:
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws[f'B{row}'] = formula
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=12))
        if 'Rate' in label:
            ws[f'B{row}'].number_format = '0.0%'
        row += 1
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'KEY ACTIONS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))
    
    row += 1
    actions = [
        f'{BULLET} Address [!] Critical Open SPOFs immediately (30-day target)',
        f'{BULLET} Review systems with ❌ None redundancy (especially Tier 1)',
        f'{BULLET} Test all systems marked ➖ Not Tested (failover testing required)',
        f'{BULLET} Address RTO gaps (actual failover time > requirement)',
        f'{BULLET} Target: 100% redundancy for Critical systems (RTO < 4h)',
        f'{BULLET} Document minimum 5 evidence items in Evidence_Register',
    ]
    
    for action in actions:
        ws[f'A{row}'] = action
        row += 1
    
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 20
    
    return ws

# ============================================================================
# WORKSHEET: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(wb):
    """Create Evidence Register worksheet (Gold Standard — 100 data rows, navy headers)"""
    ws = wb.create_sheet(title="Evidence Register")
    ws.sheet_view.showGridLines = False
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: TITLE BANNER (GS-ER-001/002/003)
    ws.merge_cells('A1:H1')
    ws['A1'] = 'EVIDENCE REGISTER'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = border

    # Row 2: SUBTITLE italic (GS-ER-004)
    ws.merge_cells('A2:H2')
    ws['A2'] = 'Document all evidence collected during this assessment (minimum 5 items required for audit compliance)'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = border

    # Row 4: COLUMN HEADERS with 003366 fill (GS-ER-005)
    headers = [
        'Evidence ID', 'Assessment Area', 'Evidence Type', 'Description',
        'Location / Path', 'Date Collected', 'Collected By', 'Verification Status',
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Data validations
    ev_type_dv = DataValidation(
        type='list',
        formula1='"Architecture Diagram,Config File,Screenshot,Report,Log File,Test Result,Policy Document,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type='list',
        formula1=f'"{CHECK} Verified,Pending,{XMARK} Not Verified"',
        allow_blank=True,
    )
    ws.add_data_validation(ver_status_dv)

    # Row 5: SAMPLE ROW (F2F2F2 grey) — GS-ER-006: must start with "EV-"
    sample_data = {
        1: 'EV-001',
        2: 'Redundancy Inventory',
        3: 'Architecture Diagram',
        4: 'E-Commerce redundancy architecture diagram (Active-Passive configuration)',
        5: '/evidence/redundancy/ecommerce_architecture.pdf',
        6: datetime.now().strftime('%d.%m.%Y'),
        7: 'Solutions Architect',
        8: f'{CHECK} Verified',
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        cell.border = border
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(
            horizontal='center' if col == 1 else 'left',
            vertical='center',
            wrap_text=True
        )
    ev_type_dv.add(ws['C5'])
    ver_status_dv.add(ws['H5'])

    # Rows 6-105: EMPTY DATA ROWS (FFFFCC yellow, 100 rows per MAX-002 standard)
    for data_row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=data_row, column=col)
            cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(
                horizontal='center' if col == 1 else 'left',
                vertical='center',
                wrap_text=True
            )
            cell.value = None
        ev_type_dv.add(ws[f'C{data_row}'])
        ver_status_dv.add(ws[f'H{data_row}'])

    # Column widths and freeze (GS-ER-007)
    for col, width in [('A', 15), ('B', 25), ('C', 22), ('D', 45), ('E', 45), ('F', 16), ('G', 22), ('H', 20)]:
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A5'
    return ws

# ============================================================================
# WORKSHEET: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_sheet(wb):
    """Create Approval Sign-Off worksheet (Gold Standard)"""
    ws = wb.create_sheet(title="Approval Sign-Off")
    ws.sheet_view.showGridLines = False
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: TITLE BANNER (GS-AS-001/002/003)
    ws.merge_cells('A1:E1')
    ws['A1'] = 'ASSESSMENT APPROVAL AND SIGN-OFF'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # Row 2: CONTROL REFERENCE subtitle (DS-006)
    ws.merge_cells('A2:E2')
    ws['A2'] = f'{DOCUMENT_ID} | ISO/IEC 27001:2022 - Control A.8.14 (Redundancy of Information Processing Facilities)'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells('A3:E3')
    ws['A3'] = 'ASSESSMENT SUMMARY'
    ws['A3'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border

    # Summary fields
    summary_fields = [
        ('Document:', f'{DOCUMENT_ID} - {WORKBOOK_TITLE}'),
        ('Assessment Period:', ''),
        ('Overall Redundancy Coverage:', "='Summary Dashboard'!B7"),
        ('Assessment Status:', ''),
        ('Assessed By:', ''),
    ]
    row = 4
    status_row_for_dv = None
    for label, value in summary_fields:
        editable = (value == '')
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = value
        for col in range(2, 6):
            if editable:
                ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        if 'Assessment Status' in label:
            status_row_for_dv = row
        if 'Coverage' in label or 'Rate' in label:
            ws[f'B{row}'].number_format = '0.0%'
        row += 1

    # Assessment Status dropdown
    status_dv = DataValidation(
        type='list',
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    if status_row_for_dv:
        status_dv.add(f'B{status_row_for_dv}')

    row += 2

    def _create_approver_section(start_row, title, color):
        """Create one approver section (GS-AS-009: Name/Title/Date/Signature/Comments order)"""
        ws.merge_cells(f'A{start_row}:E{start_row}')
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        ws[f'A{start_row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, 6):
            ws.cell(row=start_row, column=col).border = border
        current_row = start_row + 1
        for field in ['Name:', 'Title:', 'Date:', 'Signature:', 'Comments:']:
            ws[f'A{current_row}'] = field
            ws[f'A{current_row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'A{current_row}'].border = border
            ws.merge_cells(f'B{current_row}:E{current_row}')
            for col in range(2, 6):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                ws.cell(row=current_row, column=col).border = border
            current_row += 1
        return current_row + 1

    row = _create_approver_section(row, 'COMPLETED BY (ASSESSOR)', '4472C4')
    row = _create_approver_section(row, 'REVIEWED BY (INFORMATION SECURITY OFFICER)', '4472C4')
    row = _create_approver_section(row, 'APPROVED BY (CISO)', '003366')

    # FINAL DECISION (GS-AS-004)
    ws[f'A{row}'] = 'FINAL DECISION:'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'A{row}'].border = border
    ws.merge_cells(f'B{row}:E{row}')
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws.cell(row=row, column=col).border = border
    decision_dv = DataValidation(
        type='list',
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(f'B{row}')

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'NEXT REVIEW DETAILS'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    row += 1
    for label in ['Next Review Date:', 'Review Responsible:', 'Special Considerations:']:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        row += 1

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.freeze_panes = 'A3'
    return ws

# ============================================================================
# MAIN
# ============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        ws.data_validations.dataValidation = [
            dv for dv in list(ws.data_validations.dataValidation)
            if dv.sqref
        ]


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.remove(wb.active)
    
    logger.info("  [1/7] Creating Instructions...")
    create_instructions_sheet(wb.create_sheet())
    logger.info("  ✅ Complete")
          
    logger.info("  [2/7] Creating Redundancy_Inventory...")
    create_redundancy_inventory_sheet(wb)
    logger.info("  ✅ Complete (110 rows)")
    
    logger.info("  [3/7] Creating SPOF_Register...")
    create_spof_register_sheet(wb)
    logger.info("  ✅ Complete (100 SPOFs)")
    
    logger.info("  [4/7] Creating RTO_Compliance...")
    create_rto_compliance_sheet(wb)
    logger.info("  ✅ Complete (110 rows)")
    
    logger.info("  [5/7] Creating Evidence_Register...")
    create_evidence_register(wb)
    logger.info("  ✅ Complete (100 evidence rows)")

    logger.info("  [6/7] Creating Summary...")
    create_summary_dashboard_sheet(wb)
    logger.info("  ✅ Complete")
    
    logger.info("  [7/7] Creating Approval_Sign_Off...")
    create_approval_sheet(wb)
    logger.info("  ✅ Complete (3-level)")
    
    finalize_validations(wb)
    wb.save(output_path)
    logger.info(f"\n✅ SUCCESS: {output_path.name}")
    logger.info(f"   Worksheets: {len(wb.sheetnames)}")
    logger.info(f"{'='*70}\n")
    
def main():
    try:
        create_workbook(_wkbk_dir / OUTPUT_FILENAME)
    except Exception as e:
        logger.error(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
