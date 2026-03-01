#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.30.S3 - RPO/RTO Compliance Matrix Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.5.30 (Backup), A.8.14 (Redundancy), A.5.30 (ICT BC)
Assessment Domain 3 of 4: Recovery Objectives Alignment

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific RPO/RTO requirements derived from Business Impact
Analysis (BIA) and recovery capability assessment.

Key customization areas:
1. System criticality tiers and associated RPO/RTO targets
2. RPO/RTO measurement units and granularity
3. Compliance threshold definitions (meets/at risk/non-compliant)
4. Gap prioritization criteria (criticality × gap = risk score)
5. Remediation cost vs. risk tolerance thresholds

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.30/14/5.30 BC/DR Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
alignment between business-defined Recovery Point Objectives (RPO) and Recovery
Time Objectives (RTO) versus actual technical recovery capabilities.

**Purpose:**
Enables systematic assessment of recovery capability gaps against ISO 27001:2022
Controls A.5.30, A.8.14, and A.5.30 requirements, supporting evidence-based
prioritization of BC/DR investments and remediation efforts.

**Assessment Scope:**
- System inventory with business-defined RPO/RTO requirements
- Actual backup capability assessment (achievable RPO)
- Actual redundancy/failover capability assessment (achievable RTO)
- RPO compliance analysis (requirement vs. backup frequency)
- RTO compliance analysis (requirement vs. failover capability)
- Gap identification and risk scoring (criticality × gap)
- Remediation prioritization and cost-benefit analysis
- Business impact assessment for non-compliant systems
- Compensating controls documentation
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and RPO/RTO methodology
2. RPO/RTO Matrix - System recovery requirements vs. capabilities (110 rows: 10 examples + 100 data entry)
3. Evidence Register - Audit evidence tracking and documentation (100 rows)
4. Approval & Sign-Off - Multi-level stakeholder review and approval workflow

**Key Features:**
- Data validation with comprehensive dropdown lists (compliance status, gap severity)
- Conditional formatting for gap visualization and risk highlighting
- Automated compliance calculations (RPO: requirement vs. backup frequency)
- Automated compliance calculations (RTO: requirement vs. failover time)
- Automated risk scoring (criticality × gap magnitude)
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- 3-level approval workflow (Assessor -> ISO Officer -> CISO)
- Exception handling with graceful error reporting
- Professional styling without Excel repair warnings

**Integration:**
This assessment feeds into the ISMS-IMP-A.5.30.S5 BC/DR Summary Dashboard,
which consolidates data from all four BC/DR assessment domains for executive
oversight and audit readiness. Data sources from Domain 1 (Backup) and Domain 2
(Redundancy) inform this compliance analysis.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl --break-system-packages

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a530_3_rpo_rto_compliance.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a530_3_rpo_rto_compliance.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a530_3_rpo_rto_compliance.py --date 20250125

Output:
    File: ISMS_Assessment_RPO_RTO_Compliance.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and validate BIA-derived RPO/RTO requirements
    2. Inventory all business-critical systems requiring recovery
    3. Document actual backup capabilities (from Domain 1 assessment)
    4. Document actual redundancy capabilities (from Domain 2 assessment)
    5. Analyze RPO/RTO compliance gaps
    6. Prioritize gaps based on risk scoring
    7. Define remediation actions with cost-benefit analysis
    8. Document compensating controls for accepted gaps
    9. Obtain stakeholder approvals
    10. Feed results into ISMS-IMP-A.5.30.S5 BC/DR Summary Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.5.30, A.8.14, A.5.30
Assessment Domain:    3 of 4 (RPO/RTO Compliance & Gap Analysis)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.30-8.13-14: BC/DR Framework Policy (Governance)
    - ISMS-POL-A.5.30-8.13-14-S2: Information Backup Requirements (A.5.30)
    - ISMS-POL-A.5.30-8.13-14-S3: Redundancy Requirements (A.8.14)
    - ISMS-POL-A.5.30-8.13-14-S4: ICT BC Readiness Requirements (A.5.30)
    - ISMS-IMP-A.5.30-8.13-14-S1: BIA and RPO/RTO Process
    - ISMS-IMP-A.5.30.S1: Backup Inventory Assessment (Domain 1)
    - ISMS-IMP-A.5.30.S2: Redundancy Analysis Assessment (Domain 2)
    - ISMS-IMP-A.5.30.S4: BC/DR Testing Results Tracker (Domain 4)
    - ISMS-IMP-A.5.30.S5: BC/DR Summary Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.30-8.13-14-S1 specification
    - Supports comprehensive RPO/RTO compliance and gap analysis
    - Integrated with ISMS-IMP-A.5.30.S5 BC/DR Summary Dashboard
    - Includes automated risk scoring and gap prioritization

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Business Requirements Drive Technical Capabilities:**
RPO/RTO requirements must derive from Business Impact Analysis (BIA), not
IT assumptions. This assessment identifies where technical capabilities
don't meet business needs, enabling evidence-based investment decisions.

**Gap Interpretation:**
RPO/RTO gaps represent risk, not automatic compliance failures:
- Meets Requirement: Technical capability ≤ business requirement (GOOD)
- At Risk: Small gap, compensating controls may be acceptable
- Non-Compliant: Large gap requiring remediation or risk acceptance

Document business decisions to accept gaps with compensating controls.

**Risk Scoring Methodology:**
Risk = Criticality × Gap Magnitude
- Tier 1 Critical system with 8-hour gap: HIGH RISK
- Tier 4 Low system with 8-hour gap: LOW RISK

Use risk scoring to prioritize remediation investments, not to shame IT.

**RPO Gap Analysis:**
RPO gap occurs when backup frequency > RPO requirement:
- Requirement: RPO = 4 hours (backup every 4 hours)
- Actual: Daily backup (24 hours)
- Gap: 20 hours of potential data loss
- Risk: Criticality-dependent

**RTO Gap Analysis:**
RTO gap occurs when recovery time > RTO requirement:
- Requirement: RTO = 2 hours (recover within 2 hours)
- Actual: Manual rebuild = 8 hours
- Gap: 6 hours additional downtime
- Risk: Criticality-dependent

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect documented BIA results and gap remediation plans.

**Data Protection:**
Assessment workbooks contain sensitive business information including:
- System criticality classifications
- Business impact assessments
- Recovery capability gaps (risk information)

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Quarterly: Re-validate RPO/RTO requirements with business stakeholders
- Semi-annually: Assess technical capability improvements
- Annually: Complete BIA refresh and full reassessment
- Ad-hoc: After major infrastructure changes or business process changes

**Quality Assurance:**
Have business continuity managers, business process owners, and IT leadership
validate assessments before using results for investment decisions.

**Regulatory Alignment:**
Ensure RPO/RTO requirements align with applicable regulatory requirements:
- Finance: DORA recovery time objectives for critical services
- Healthcare: HIPAA contingency planning requirements
- Critical Infrastructure: Sector-specific recovery mandates

**Cost-Benefit Analysis:**
Not all gaps require remediation. Document business-driven decisions:
- Remediation cost < risk exposure: REMEDIATE
- Remediation cost > risk exposure: ACCEPT RISK (with appropriate approval)
- Compensating controls available: DOCUMENT AND MONITOR

This assessment provides data for informed decision-making, not blame assignment.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from datetime import datetime, timedelta
import sys
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
WORKBOOK_TITLE = "RPO/RTO Compliance Matrix"
WORKBOOK_NAME = "RPO RTO Compliance"
VERSION = "1.0"
CONTROLS = "A.5.30 (Information Backup)"


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠  Warning sign
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.30.S3"
CONTROL_ID   = "A.5.30"
CONTROL_NAME = "Information Backup"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)
_assessment_dir = _wkbk_dir / "Assessment"
if _assessment_dir.exists():
    _wkbk_dir = _assessment_dir

# Color scheme (consistent with reference implementations)
HEADER_FILL = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
COLUMN_HEADER_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

# Fonts
HEADER_FONT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
SUBHEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
BOLD_FONT = Font(name='Calibri', size=10, bold=True)
NORMAL_FONT = Font(name='Calibri', size=10)

# Borders
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def apply_style(cell, font=None, fill=None, alignment=None, border=None):
    """Apply multiple styles to a cell"""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border

def create_base_validations(ws):
    """
    Create comprehensive data validation objects for dropdowns.
    
    Returns dict of validation objects to be applied to cells.
    All validations are added to worksheet once, then applied to multiple cells.
    """
    validations = {
        'criticality': DataValidation(
            type="list",
            formula1='"Tier 1 - Critical,Tier 2 - Important,Tier 3 - Standard,Tier 4 - Low"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Compliant,⚠️ Partial,❌ Non-Compliant,❓ Unknown"',
            allow_blank=False
        ),
        'priority': DataValidation(
            type="list",
            formula1='"[!] Critical,[~] High,[.] Medium,[ ] Low"',
            allow_blank=False
        ),
        'gap_type': DataValidation(
            type="list",
            formula1='"RPO Gap,RTO Gap,Both RPO & RTO,Testing Gap,Documentation Gap"',
            allow_blank=False
        ),
        'gap_status': DataValidation(
            type="list",
            formula1=f'"Open,In Progress,{CHECK} Closed,Risk Accepted"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Config File,Screenshot,Report,Log File,Test Result,Policy Document,BIA Report,Contract,Diagram,Other"',
            allow_blank=False
        ),
        'verification_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Verified,⏳ Pending,❌ Not Verified"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1='"Draft,Under Review,Final,Requires Remediation"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected,Requires Rework"',
            allow_blank=False
        ),
        'recommendation': DataValidation(
            type="list",
            formula1='"Approve,Approve with Conditions,Reject,Require Rework"',
            allow_blank=False
        ),
    }
    
    return validations

# ============================================================================
# WORKSHEET: INSTRUCTIONS & LEGEND
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Populate the System Inventory worksheet from Business Impact Analysis (BIA) results.', '2. Populate the Capability Assessment worksheet with TESTED technical recovery capabilities.', '3. Review the Compliance Matrix (formulas automatically compare requirements vs capabilities).', '4. Document identified gaps in the Gap Analysis worksheet with remediation plans.', '5. Collect evidence in the Evidence Register (minimum 5 items required for audit).', '6. Review the Summary Dashboard for overall compliance metrics.', '7. Complete the Approval Sign-Off worksheet to obtain formal sign-off.', '8. Re-assess quarterly to track improvement and maintain compliance.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 22

    _EVIDENCE = ['✓ Business Impact Analysis (BIA) report with RPO/RTO requirements for each system', '✓ Backup restore test records with date, system, and actual restore time', '✓ Failover test records with date, system, and actual failover time', '✓ Gap analysis report showing RPO/RTO gaps and risk scores', '✓ Remediation plans for identified gaps with target dates', '✓ System criticality register from BIA process', '✓ Testing schedules and historical test logs', '✓ Compliance calculation worksheets and summary reports']

    # Status Legend — row position tracks after instructions
    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Acceptable Evidence section
    _ev_row = _leg_row + 7
    ws[f"A{_ev_row}"] = "Acceptable Evidence (examples)"
    ws[f"A{_ev_row}"].font = Font(name="Calibri", size=12, bold=True)
    for _ev in _EVIDENCE:
        _ev_row += 1
        ws[f"A{_ev_row}"] = _ev

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_summary_dashboard_sheet(wb):
    """Create Summary Dashboard worksheet (placed as 2nd sheet)"""
    ws = wb.create_sheet(title="Summary Dashboard")
    ws.sheet_view.showGridLines = False

    # Title — DS-002: fill must be HEADER_FILL (003366) with white font
    ws.merge_cells('A1:E1')
    ws['A1'] = 'RPO/RTO COMPLIANCE \u2014 SUMMARY DASHBOARD'
    apply_style(ws['A1'], font=Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
                fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:E2')
    ws['A2'] = f'Assessment ID: {DOCUMENT_ID} | Assessment Date: [enter date]'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(italic=True, size=10)

    # Overall Compliance Metrics
    row = 4
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'OVERALL RPO/RTO COMPLIANCE'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))

    row += 1
    compliance_metrics = [
        ('Total Systems Assessed:', "=COUNTA('System Inventory'!A16:A65)"),
        ('Systems with Full Compliance (RPO & RTO):', f"=COUNTIF('Compliance Matrix'!E16:E65,\"{CHECK}*\")"),
        ('Systems with Partial Compliance:', f"=COUNTIF('Compliance Matrix'!E16:E65,\"{WARNING}*\")"),
        ('Non-Compliant Systems:', f"=COUNTIF('Compliance Matrix'!E16:E65,\"{XMARK}*\")"),
        ('Unknown Status (Not Tested):', "=COUNTIF('Compliance Matrix'!E16:E65,\"❓*\")"),
        ('Overall Compliance Rate:', '=IF(B5>0,B6/B5,0)'),
    ]

    for label, formula in compliance_metrics:
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws[f'B{row}'] = formula
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=12))
        if 'Rate' in label:
            ws[f'B{row}'].number_format = '0.0%'
        row += 1

    # Criticality Breakdown
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'COMPLIANCE BY CRITICALITY TIER'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))

    row += 1
    criticality_metrics = [
        ('Tier 1 - Critical Systems:', "=COUNTIF('System Inventory'!C16:C65,\"Tier 1*\")"),
        ('Tier 1 - Full Compliance:', f"=SUMPRODUCT(('System Inventory'!C16:C65=\"Tier 1 - Critical\")*('Compliance Matrix'!E16:E65=\"{CHECK} Full Compliance\"))"),
        ('Tier 1 - Compliance Rate:', '=IF(B13>0,B14/B13,0)'),
        ('',  ''),
        ('Tier 2 - Important Systems:', "=COUNTIF('System Inventory'!C16:C65,\"Tier 2*\")"),
        ('Tier 2 - Full Compliance:', f"=SUMPRODUCT(('System Inventory'!C16:C65=\"Tier 2 - Important\")*('Compliance Matrix'!E16:E65=\"{CHECK} Full Compliance\"))"),
        ('Tier 2 - Compliance Rate:', '=IF(B17>0,B18/B17,0)'),
    ]

    for label, formula in criticality_metrics:
        if label:
            ws[f'A{row}'] = label
            apply_style(ws[f'A{row}'], font=BOLD_FONT)
            ws[f'B{row}'] = formula
            apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
            if 'Rate' in label:
                ws[f'B{row}'].number_format = '0.0%'
        row += 1

    # Gap Analysis Summary
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'GAP ANALYSIS SUMMARY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))

    row += 1
    gap_metrics = [
        ('Total Gaps Identified:', "=COUNTA('Gap Analysis'!A16:A65)"),
        ('[!] Critical Priority Gaps:', "=COUNTIF('Gap Analysis'!F16:F65,\"[!]*\")"),
        ('[~] High Priority Gaps:', "=COUNTIF('Gap Analysis'!F16:F65,\"[~]*\")"),
        ('[.] Medium Priority Gaps:', "=COUNTIF('Gap Analysis'!F16:F65,\"[.]*\")"),
        ('[ ] Low Priority Gaps:', "=COUNTIF('Gap Analysis'!F16:F65,\"[ ]*\")"),
        ('',  ''),
        ('Open Gaps:', "=COUNTIF('Gap Analysis'!H16:H65,\"Open*\")"),
        ('In Progress Gaps:', "=COUNTIF('Gap Analysis'!H16:H65,\"In Progress*\")"),
        ('Closed Gaps:', f"=COUNTIF('Gap Analysis'!H16:H65,\"{CHECK}*\")"),
        ('',  ''),
        ('Average Risk Score:', "=IFERROR(AVERAGE('Gap Analysis'!E16:E65),0)"),
    ]

    for label, formula in gap_metrics:
        if label:
            ws[f'A{row}'] = label
            apply_style(ws[f'A{row}'], font=BOLD_FONT)
            ws[f'B{row}'] = formula
            apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
            if 'Score' in label:
                ws[f'B{row}'].number_format = '0.0'
        row += 1

    # Evidence Compliance
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'EVIDENCE & APPROVAL STATUS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))

    row += 1
    evidence_metrics = [
        ('Evidence Items Collected:', "=COUNTA('Evidence Register'!A6:A105)"),
        ('Verified Evidence:', f"=COUNTIF('Evidence Register'!H6:H105,\"{CHECK}*\")"),
        ('Minimum Evidence Required:', '5'),
        ('Evidence Compliance:', f'=IF(B40>=B42,"{CHECK} Sufficient","{XMARK} Insufficient")'),
        ('',  ''),
        ('Assessment Status:', "='Approval Sign-Off'!B7"),
        ('Level 1 - Assessor Completed:', f"=IF('Approval Sign-Off'!B11<>\"\",\"{CHECK} Complete\",\"Pending\")"),
        ('Level 2 - ISO Review:', f"=IF('Approval Sign-Off'!B18<>\"\",\"{CHECK} Complete\",\"Pending\")"),
        ('Level 3 - CISO Approval:', f"=IF('Approval Sign-Off'!B25<>\"\",\"{CHECK} Complete\",\"Pending\")"),
    ]

    for label, formula in evidence_metrics:
        if label:
            ws[f'A{row}'] = label
            apply_style(ws[f'A{row}'], font=BOLD_FONT)
            ws[f'B{row}'] = formula
            if label == 'Minimum Evidence Required:':
                apply_style(ws[f'B{row}'], font=Font(bold=True, size=11, color='C00000'))
            else:
                apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1

    # Priority Actions
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'PRIORITY ACTIONS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))

    row += 1
    priority_actions = [
        '1. Address all [!] Critical priority gaps immediately (target: 90 days)',
        '2. Test all systems marked unknown (capability assessment needed)',
        '3. Focus on Tier 1 systems first (business-critical)',
        '4. Close RPO gaps: Increase backup frequency OR implement replication',
        '5. Close RTO gaps: Implement redundancy OR optimise restore procedures',
        '6. Collect minimum 5 evidence items in the Evidence Register',
        '7. Complete all 3 approval levels (Assessor to ISO to CISO)',
        '8. Track remediation progress in the Gap Analysis worksheet',
        '9. Reassess quarterly to measure continuous improvement',
    ]

    for action in priority_actions:
        ws[f'A{row}'] = action
        ws[f'A{row}'].font = NORMAL_FONT
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 25
    for col in ['C', 'D', 'E']:
        ws.column_dimensions[col].width = 15

    return ws

# ============================================================================
# WORKSHEET: SYSTEM INVENTORY (BIA RESULTS)
# ============================================================================

def create_system_inventory_sheet(wb):
    """Create System Inventory worksheet (BIA results - business requirements)"""
    ws = wb.create_sheet(title="System Inventory")
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:G1')
    ws['A1'] = 'SYSTEM INVENTORY - BIA RESULTS'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:G2')
    ws['A2'] = 'Source: Business Impact Analysis (BIA) - RPO/RTO requirements from business stakeholders'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'System Name',
        'Business Process',
        'Criticality Tier',
        'MTPD (hours)',
        'RPO Requirement (hours)',
        'RTO Requirement (hours)',
        'Business Justification'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    # Apply data validations for 61 rows (10 examples + 1 grey sample + 50 empty)
    _grey_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row += 1
    for i in range(61):
        current_row = row + i
        # Criticality dropdown
        validations['criticality'].add(f'C{current_row}')

        # Input fills for user entry
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            if i < 10:
                apply_style(ws[f'{col}{current_row}'], border=THIN_BORDER)
            elif i == 10:  # Grey sample row (GS-MAX-003)
                apply_style(ws[f'{col}{current_row}'], fill=_grey_fill, border=THIN_BORDER)
            else:
                apply_style(ws[f'{col}{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)
        if i == 10:  # Sample data for grey row
            ws[f'A{current_row}'] = 'New System (Example)'
            ws[f'B{current_row}'] = 'Core Business Process'
            ws[f'C{current_row}'] = 'Tier 2 - Important'
            ws[f'D{current_row}'] = 24
            ws[f'E{current_row}'] = 4
            ws[f'F{current_row}'] = 8
            ws[f'G{current_row}'] = 'Sample — enter BIA-sourced requirements here'
    
    # Grey sample row at row 15 (F2F2F2) — MAX-003 fix
    _grey = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _si_sample_row = row + 10  # first blank row after 10 examples
    _si_sample = ['Sample System', 'Business Process', 'Tier 1 - Critical', 4, 1, 2,
                  'Business justification for RPO/RTO requirements']
    for _ci, _val in enumerate(_si_sample, start=1):
        _c = ws.cell(row=_si_sample_row, column=_ci, value=_val)
        _c.fill = _grey
        _c.border = THIN_BORDER
        if _ci == 3: validations['criticality'].add(_c)

    # Write sample row data (row 15 = grey F2F2F2 guide row)
    _sf = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _sv = ['Example: Application Server', 'Online Services', 'Tier 1 - Critical',
           4, 1, 2, 'Enter business justification for RPO/RTO requirements here']
    for _ci, _v in enumerate(_sv, start=1):
        _c = ws.cell(row=15, column=_ci, value=_v)
        _c.fill = _sf
        _c.border = THIN_BORDER
        if _ci == 7: _c.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Info notes
    info_row = 67
    ws.merge_cells(f'A{info_row}:G{info_row}')
    ws[f'A{info_row}'] = 'DATA SOURCE: Business Impact Analysis (BIA) - Requirements are business-driven, not IT-driven'
    ws[f'A{info_row}'].font = Font(italic=True, size=10, color='666666')
    
    info_row += 1
    ws.merge_cells(f'A{info_row}:G{info_row}')
    ws[f'A{info_row}'] = 'MTPD = Maximum Tolerable Period of Disruption | RPO = Recovery Point Objective | RTO = Recovery Time Objective'
    ws[f'A{info_row}'].font = Font(italic=True, size=10, color='666666')
    
    # Column widths
    widths = [25, 25, 18, 15, 22, 22, 45]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Lazy-add: only register DVs that have cells assigned (fixes empty-sqref XML error)
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)
    
    ws.freeze_panes = 'A4'
    
    return ws

# ============================================================================
# WORKSHEET: CAPABILITY ASSESSMENT
# ============================================================================

def create_capability_assessment_sheet(wb):
    """Create Capability Assessment worksheet (technical capabilities from testing)"""
    ws = wb.create_sheet(title="Capability Assessment")
    ws.sheet_view.showGridLines = False
    # Note: no DVs used in this formula-driven sheet — do NOT call create_base_validations
    # (empty DVs cause HRESULT 0x8000ffff XML repair error in Excel)

    # Header
    ws.merge_cells('A1:G1')
    ws['A1'] = 'CAPABILITY ASSESSMENT - TESTED TECHNICAL CAPABILITIES'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:G2')
    ws['A2'] = 'Source: Testing Results - Actual measured recovery times (restore tests, failover tests)'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))

    
    # Column headers
    row = 4
    headers = [
        'System Name',
        'Backup Frequency (hours)',
        'Restore Time Tested (hours)',
        'Failover Time Tested (hours)',
        'RPO Capability (hours)',
        'RTO Capability (hours)',
        'Test Notes / Observations'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    # Apply formulas and input fills for 61 rows (10 examples + 1 grey sample + 50 empty)
    _grey_cap = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row = 5
    for i in range(61):
        current_row = row + i

        # System Name
        if i < 10:
            apply_style(ws[f'A{current_row}'], border=THIN_BORDER)
        elif i == 10:  # Grey sample row (GS-MAX-003)
            apply_style(ws[f'A{current_row}'], fill=_grey_cap, border=THIN_BORDER)
            ws[f'A{current_row}'] = 'Example System (Sample)'
        else:
            apply_style(ws[f'A{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)

        # Backup Frequency, Restore Time, Failover Time (user input)
        for col in ['B', 'C', 'D', 'G']:
            if i < 10:
                apply_style(ws[f'{col}{current_row}'], border=THIN_BORDER)
            elif i == 10:
                apply_style(ws[f'{col}{current_row}'], fill=_grey_cap, border=THIN_BORDER)
            else:
                apply_style(ws[f'{col}{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)
        
        # E: RPO Capability = Backup Frequency
        ws[f'E{current_row}'] = f'=IF(ISNUMBER(B{current_row}),B{current_row},"")'
        apply_style(ws[f'E{current_row}'], border=THIN_BORDER,
                   font=Font(italic=True, color='666666'))
        
        # F: RTO Capability = MIN(Restore Time, Failover Time) - whichever is faster
        ws[f'F{current_row}'] = f'=IF(AND(ISNUMBER(C{current_row}),ISNUMBER(D{current_row})),MIN(C{current_row},D{current_row}),IF(ISNUMBER(C{current_row}),C{current_row},IF(ISNUMBER(D{current_row}),D{current_row},"")))'
        apply_style(ws[f'F{current_row}'], border=THIN_BORDER,
                   font=Font(italic=True, color='666666'))
    
    # Sample row 15 data (F2F2F2 grey — guide row)
    _sf = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _sv = ['Example: Application Server', 1, 4, '', '', '', 'Enter tested recovery times and observations']
    for _ci, _v in enumerate(_sv, start=1):
        if _ci not in [5, 6]:
            _c = ws.cell(row=15, column=_ci, value=_v)
            _c.fill = _sf
            _c.border = THIN_BORDER
            if _ci == 7: _c.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Info notes
    info_row = 67
    ws.merge_cells(f'A{info_row}:G{info_row}')
    ws[f'A{info_row}'] = 'DATA SOURCE: Testing Results - CRITICAL: Capabilities must be TESTED, not assumed. Untested = Unknown compliance.'
    ws[f'A{info_row}'].font = Font(italic=True, size=10, bold=True, color='C00000')
    
    info_row += 1
    ws.merge_cells(f'A{info_row}:G{info_row}')
    ws[f'A{info_row}'] = 'RPO Capability = Backup Frequency | RTO Capability = MIN(Restore Time, Failover Time) - formulas auto-calculate'
    ws[f'A{info_row}'].font = Font(italic=True, size=10, color='666666')
    
    # Column widths
    widths = [25, 24, 26, 28, 22, 22, 35]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = 'A4'

    return ws

# ============================================================================
# WORKSHEET: COMPLIANCE MATRIX
# ============================================================================

def create_compliance_matrix_sheet(wb):
    """Create Compliance Matrix worksheet (automated requirement vs capability comparison)"""
    ws = wb.create_sheet(title="Compliance Matrix")
    ws.sheet_view.showGridLines = False
    # Note: no DVs used in this formula-driven sheet — do NOT call create_base_validations
    # (empty DVs cause HRESULT 0x8000ffff XML repair error in Excel)

    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = 'RPO/RTO COMPLIANCE MATRIX'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:H2')
    ws['A2'] = "Automatic comparison: Business Requirements (System Inventory) vs Technical Capabilities (Capability Assessment)"
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'System Name',
        'Criticality',
        'RPO: Req vs Cap',
        'RTO: Req vs Cap',
        'Overall Compliance',
        'Priority',
        'Gap Summary',
        'Next Action Required'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    # Apply formulas for 61 rows (10 examples + 1 grey sample + 50 empty)
    _grey_cm = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row = 5
    for i in range(61):
        current_row = row + i
        _is_sample = (i == 10)
        _is_empty = (i > 10)

        # A: System Name (from System Inventory)
        ws[f'A{current_row}'] = f"='System Inventory'!A{current_row}"
        apply_style(ws[f'A{current_row}'],
                   fill=_grey_cm if _is_sample else None, border=THIN_BORDER)

        # B: Criticality (from System Inventory)
        ws[f'B{current_row}'] = f"='System Inventory'!C{current_row}"
        apply_style(ws[f'B{current_row}'],
                   fill=_grey_cm if _is_sample else None, border=THIN_BORDER)

        # C: RPO Compliance Status
        # Compliant if RPO Capability <= RPO Requirement
        ws[f'C{current_row}'] = f"=IF(A{current_row}=\"\",\"\",IF(OR('System Inventory'!E{current_row}=\"\",'Capability Assessment'!E{current_row}=\"\"),\"❓ Unknown\",IF('Capability Assessment'!E{current_row}<='System Inventory'!E{current_row},\"{CHECK} Compliant\",\"{XMARK} Non-Compliant\")))"
        apply_style(ws[f'C{current_row}'],
                   fill=_grey_cm if _is_sample else None, border=THIN_BORDER)

        # D: RTO Compliance Status
        # Compliant if RTO Capability <= RTO Requirement
        ws[f'D{current_row}'] = f"=IF(A{current_row}=\"\",\"\",IF(OR('System Inventory'!F{current_row}=\"\",'Capability Assessment'!F{current_row}=\"\"),\"❓ Unknown\",IF('Capability Assessment'!F{current_row}<='System Inventory'!F{current_row},\"{CHECK} Compliant\",\"{XMARK} Non-Compliant\")))"
        apply_style(ws[f'D{current_row}'],
                   fill=_grey_cm if _is_sample else None, border=THIN_BORDER)

        # E: Overall Compliance
        # Both RPO and RTO must be compliant for full compliance
        ws[f'E{current_row}'] = f"=IF(A{current_row}=\"\",\"\",IF(OR(C{current_row}=\"❓ Unknown\",D{current_row}=\"❓ Unknown\"),\"❓ Unknown\",IF(AND(C{current_row}=\"{CHECK} Compliant\",D{current_row}=\"{CHECK} Compliant\"),\"{CHECK} Full Compliance\",IF(OR(C{current_row}=\"{CHECK} Compliant\",D{current_row}=\"{CHECK} Compliant\"),\"{WARNING} Partial Compliance\",\"{XMARK} Non-Compliant\"))))"
        apply_style(ws[f'E{current_row}'],
                   fill=_grey_cm if _is_sample else None, border=THIN_BORDER)

        # F: Priority (Critical = Tier 1 + Non-Compliant)
        ws[f'F{current_row}'] = f"=IF(A{current_row}=\"\",\"\",IF(AND(B{current_row}=\"Tier 1 - Critical\",E{current_row}=\"{XMARK} Non-Compliant\"),\"[!] Critical\",IF(B{current_row}=\"Tier 1 - Critical\",\"[~] High\",IF(E{current_row}=\"{XMARK} Non-Compliant\",\"[~] High\",IF(E{current_row}=\"{WARNING} Partial Compliance\",\"[.] Medium\",\"[ ] Low\")))))"
        apply_style(ws[f'F{current_row}'],
                   fill=_grey_cm if _is_sample else None, border=THIN_BORDER)

        # G: Gap Summary
        ws[f'G{current_row}'] = f"=IF(A{current_row}=\"\",\"\",IF(C{current_row}=\"{XMARK} Non-Compliant\",IF(D{current_row}=\"{XMARK} Non-Compliant\",\"Both RPO & RTO gaps\",\"RPO gap only\"),IF(D{current_row}=\"{XMARK} Non-Compliant\",\"RTO gap only\",IF(E{current_row}=\"❓ Unknown\",\"Testing required\",\"No gaps\"))))"
        apply_style(ws[f'G{current_row}'],
                   fill=_grey_cm if _is_sample else None,
                   border=THIN_BORDER, alignment=Alignment(wrap_text=True))

        # H: Next Action (manual input)
        if _is_sample:
            apply_style(ws[f'H{current_row}'], fill=_grey_cm, border=THIN_BORDER,
                       alignment=Alignment(wrap_text=True))
            ws[f'H{current_row}'] = 'Sample — enter next remediation action here'
        elif _is_empty:
            apply_style(ws[f'H{current_row}'], fill=INPUT_FILL, border=THIN_BORDER,
                       alignment=Alignment(wrap_text=True))
        else:
            apply_style(ws[f'H{current_row}'], border=THIN_BORDER,
                       alignment=Alignment(wrap_text=True))

    # Example next actions for first 10 rows (matching example systems)
    next_actions = [
        'Monitor continuously, maintain current state',
        'Monitor continuously, critical system',
        'Consider more frequent backups (currently 4hr, need 4hr RTO)',
        'Implement hot standby for faster recovery',
        'Acceptable compliance, monitor quarterly',
        'Acceptable compliance, monitor quarterly',
        'Acceptable for Tier 3, no action required',
        'Acceptable for Tier 3, no action required',
        'Acceptable for Tier 4, no action required',
        'Acceptable for Tier 4, no action required',
    ]

    row = 5
    for idx, action in enumerate(next_actions, start=row):
        ws[f'H{idx}'] = action
        ws[f'H{idx}'].alignment = Alignment(wrap_text=True)

    # Summary metrics (data rows 5-65: 61 rows total)
    summary_row = 68
    ws.merge_cells(f'A{summary_row}:H{summary_row}')
    ws[f'A{summary_row}'] = 'COMPLIANCE SUMMARY METRICS'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))

    summary_row += 1
    metrics = [
        ('Total Systems:', f'=COUNTA(A6:A65)'),
        ('Full Compliance (✅):', f'=COUNTIF(E6:E65,"{CHECK}*")'),
        ('Partial Compliance (⚠️):', f'=COUNTIF(E6:E65,"{WARNING}*")'),
        ('Non-Compliant (❌):', f'=COUNTIF(E6:E65,"{XMARK}*")'),
        ('Unknown (❓):', f'=COUNTIF(E6:E65,"❓*")'),
        ('Overall Compliance Rate:', f'=IF(B{summary_row}>0,B{summary_row+1}/B{summary_row},0)'),
    ]
    
    for label, formula in metrics:
        ws[f'A{summary_row}'] = label
        apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
        ws[f'B{summary_row}'] = formula
        apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
        if 'Rate' in label:
            ws[f'B{summary_row}'].number_format = '0.0%'
        summary_row += 1
    
    # Column widths
    widths = [25, 18, 18, 18, 22, 15, 25, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = 'A4'
    
    return ws

# ============================================================================
# WORKSHEET: GAP ANALYSIS
# ============================================================================

def create_gap_analysis_sheet(wb):
    """Create Gap Analysis worksheet with remediation tracking"""
    ws = wb.create_sheet(title="Gap Analysis")
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = 'GAP ANALYSIS & REMEDIATION TRACKING'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:H2')
    ws['A2'] = 'Document identified gaps with risk scoring and remediation plans'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'Gap ID',
        'System',
        'Gap Type',
        'Gap Description',
        'Risk Score (1-10)',
        'Priority',
        'Remediation Plan',
        'Status'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    # Apply data validations and styling for 61 rows (10 examples + 1 grey sample + 50 empty)
    _grey_gap = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row = 5
    for i in range(61):
        current_row = row + i

        # Gap ID — example rows (i<10) left empty (IDs added by examples section below);
        # grey sample (i=10) gets a placeholder (GS-MAX-003)
        if i < 10:
            apply_style(ws[f'A{current_row}'], font=Font(bold=True, size=9), border=THIN_BORDER)
        elif i == 10:  # Grey sample row (GS-MAX-003)
            ws[f'A{current_row}'] = 'GAP-NNN'
            apply_style(ws[f'A{current_row}'], fill=_grey_gap,
                       font=Font(bold=True, size=9), border=THIN_BORDER)
        else:
            apply_style(ws[f'A{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)

        # System, Gap Description, Remediation Plan (user input)
        for col in ['B', 'D', 'G']:
            if i < 10:
                apply_style(ws[f'{col}{current_row}'], border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
            elif i == 10:
                apply_style(ws[f'{col}{current_row}'], fill=_grey_gap, border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
            else:
                apply_style(ws[f'{col}{current_row}'], fill=INPUT_FILL, border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))

        # Gap Type dropdown
        validations['gap_type'].add(f'C{current_row}')
        apply_style(ws[f'C{current_row}'],
                   fill=_grey_gap if i == 10 else (INPUT_FILL if i > 10 else None),
                   border=THIN_BORDER)

        # Risk Score (user input, numeric 1-10)
        if i < 10:
            apply_style(ws[f'E{current_row}'], border=THIN_BORDER)
        elif i == 10:
            apply_style(ws[f'E{current_row}'], fill=_grey_gap, border=THIN_BORDER)
            ws[f'E{current_row}'] = 5
        else:
            apply_style(ws[f'E{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)

        # Priority dropdown
        validations['priority'].add(f'F{current_row}')
        apply_style(ws[f'F{current_row}'],
                   fill=_grey_gap if i == 10 else None, border=THIN_BORDER)

        # Status dropdown
        validations['gap_status'].add(f'H{current_row}')
        apply_style(ws[f'H{current_row}'],
                   fill=_grey_gap if i == 10 else None, border=THIN_BORDER)

        if i == 10:  # Additional sample data
            ws[f'B{current_row}'] = 'New System (Example)'
            ws[f'C{current_row}'] = 'RPO Gap'
            ws[f'D{current_row}'] = 'RPO requirement not met — enter actual gap description here'
            ws[f'F{current_row}'] = '[~] High'
            ws[f'G{current_row}'] = 'Sample — enter remediation plan here'
            ws[f'H{current_row}'] = 'Open'
    
    # Grey sample row at row 15 (F2F2F2) — MAX-003 fix
    _grey_ga = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _ga_sample_row = 5 + 10  # row 15
    _ga_sample = ['GAP-NNN', 'System Name', 'RTO Gap',
                  'Gap description — actual vs required RPO/RTO', 7, '[~] High',
                  'Remediation plan and target timeline', 'In Progress']
    for _ci, _val in enumerate(_ga_sample, start=1):
        _c = ws.cell(row=_ga_sample_row, column=_ci, value=_val)
        _c.fill = _grey_ga
        _c.border = THIN_BORDER

    # Summary metrics
    summary_row = 67
    ws.merge_cells(f'A{summary_row}:H{summary_row}')
    ws[f'A{summary_row}'] = 'GAP SUMMARY METRICS'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    summary_row += 1
    metrics = [
        ('Total Gaps Identified:', f'=COUNTA(B5:B65)'),
        ('[!] Critical Priority:', f'=COUNTIF(F5:F65,"[!]*")'),
        ('[~] High Priority:', f'=COUNTIF(F5:F65,"[~]*")'),
        ('[.] Medium Priority:', f'=COUNTIF(F5:F65,"[.]*")'),
        ('[ ] Low Priority:', f'=COUNTIF(F5:F65,"[ ]*")'),
        ('', ''),
        ('Open Gaps:', f'=COUNTIF(H5:H65,"Open*")'),
        ('In Progress:', f'=COUNTIF(H5:H65,"In Progress*")'),
        (f'{CHECK} Closed Gaps:', f'=COUNTIF(H5:H65,"{CHECK}*")'),
        ('Risk Accepted:', f'=COUNTIF(H5:H65,"Risk Accepted*")'),
        ('', ''),
        ('Average Risk Score:', f'=IFERROR(AVERAGE(E5:E114),0)'),
        ('Max Risk Score:', f'=IFERROR(MAX(E5:E114),0)'),
    ]
    
    for label, formula in metrics:
        if label:
            ws[f'A{summary_row}'] = label
            apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
            ws[f'B{summary_row}'] = formula
            apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
            if 'Score' in label:
                ws[f'B{summary_row}'].number_format = '0.0'
        summary_row += 1
    
    # Column widths
    widths = [12, 25, 18, 40, 18, 15, 40, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Lazy-add: only register DVs that have cells assigned (fixes empty-sqref XML error)
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)
    
    ws.freeze_panes = 'A4'
    
    return ws

# ============================================================================
# WORKSHEET: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(wb):
    """Create Evidence Register worksheet (Gold Standard — 100 data rows, navy headers)"""
    ws = wb.create_sheet(title="Evidence Register")
    ws.sheet_view.showGridLines = False
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: TITLE BANNER (GS-ER-001/002/003)
    ws.merge_cells('A1:H1')
    ws['A1'] = 'EVIDENCE REGISTER'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = border

    # Row 2: SUBTITLE italic (GS-ER-004)
    ws.merge_cells('A2:H2')
    ws['A2'] = 'Document all evidence collected during this assessment (minimum 5 items required for audit compliance)'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 9):
        ws.cell(row=2, column=col).border = border

    # Row 3: intentionally empty (visual separator)

    # Row 4: COLUMN HEADERS with 003366 fill (GS-ER-005)
    headers = [
        'Evidence ID', 'Assessment Area', 'Evidence Type', 'Description',
        'Location / Path', 'Date Collected', 'Collected By', 'Verification Status',
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Data validations
    ev_type_dv = DataValidation(
        type='list',
        formula1='"Config File,Screenshot,Report,Log File,Test Result,Policy Document,BIA Report,Diagram,Other"',
        allow_blank=True,
    )
    ws.add_data_validation(ev_type_dv)

    ver_status_dv = DataValidation(
        type='list',
        formula1=f'"{CHECK} Verified,Pending,{XMARK} Not Verified"',
        allow_blank=True,
    )
    ws.add_data_validation(ver_status_dv)

    # Row 5: SAMPLE ROW (F2F2F2 grey) — GS-ER-006: must start with "EV-"
    sample_data = {
        1: 'EV-001',
        2: 'Compliance Matrix',
        3: 'BIA Report',
        4: 'Business Impact Analysis report with RPO/RTO requirements for all systems',
        5: '/evidence/bia/BIA_Report_Q1_2024.pdf',
        6: datetime.now().strftime('%d.%m.%Y'),
        7: 'BC/DR Coordinator',
        8: f'{CHECK} Verified',
    }
    for col, value in sample_data.items():
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        cell.border = border
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(
            horizontal='center' if col == 1 else 'left',
            vertical='center',
            wrap_text=True
        )
    ev_type_dv.add(ws['C5'])
    ver_status_dv.add(ws['H5'])

    # Rows 6-105: EMPTY DATA ROWS (FFFFCC yellow, 100 rows)
    for data_row in range(6, 106):
        for col in range(1, 9):
            cell = ws.cell(row=data_row, column=col)
            cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(
                horizontal='center' if col == 1 else 'left',
                vertical='center',
                wrap_text=True
            )
            cell.value = None
        ev_type_dv.add(ws[f'C{data_row}'])
        ver_status_dv.add(ws[f'H{data_row}'])

    # Column widths and freeze (GS-ER-007)
    for col, width in [('A', 15), ('B', 25), ('C', 22), ('D', 45), ('E', 45), ('F', 16), ('G', 22), ('H', 20)]:
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A5'

    return ws

# ============================================================================
# WORKSHEET: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_sheet(wb):
    """Create Approval Sign-Off worksheet (Gold Standard)"""
    ws = wb.create_sheet(title="Approval Sign-Off")
    ws.sheet_view.showGridLines = False
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: TITLE BANNER (GS-AS-001/002/003: A1:E1 merge, "AND" not "&", height 35)
    ws.merge_cells('A1:E1')
    ws['A1'] = 'ASSESSMENT APPROVAL AND SIGN-OFF'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # Row 2: CONTROL REFERENCE subtitle (DS-006)
    ws.merge_cells('A2:E2')
    ws['A2'] = f'{DOCUMENT_ID} | ISO/IEC 27001:2022 - Control A.5.30 (ICT Readiness for Business Continuity)'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells('A3:E3')
    ws['A3'] = 'ASSESSMENT SUMMARY'
    ws['A3'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border

    # Summary fields
    summary_fields = [
        ('Document:', f'{DOCUMENT_ID} - {WORKBOOK_TITLE}'),
        ('Assessment Period:', ''),
        ('Overall RPO/RTO Compliance Rate:', "='Summary Dashboard'!B10"),
        ('Assessment Status:', ''),
        ('Assessed By:', ''),
    ]
    row = 4
    status_row_for_dv = None
    for label, value in summary_fields:
        editable = (value == '')
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = value
        for col in range(2, 6):
            if editable:
                ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        if 'Assessment Status' in label:
            status_row_for_dv = row
        if 'Rate' in label or 'Compliance' in label:
            ws[f'B{row}'].number_format = '0.0%'
        row += 1

    # Assessment Status dropdown
    status_dv = DataValidation(
        type='list',
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    if status_row_for_dv:
        status_dv.add(f'B{status_row_for_dv}')

    row += 2  # Gap before first approver

    def _create_approver_section(start_row, title, color):
        """Create one approver section (GS-AS-009: Name/Title/Date/Signature/Comments order)"""
        ws.merge_cells(f'A{start_row}:E{start_row}')
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        ws[f'A{start_row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, 6):
            ws.cell(row=start_row, column=col).border = border
        current_row = start_row + 1
        for field in ['Name:', 'Title:', 'Date:', 'Signature:', 'Comments:']:
            ws[f'A{current_row}'] = field
            ws[f'A{current_row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'A{current_row}'].border = border
            ws.merge_cells(f'B{current_row}:E{current_row}')
            for col in range(2, 6):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                ws.cell(row=current_row, column=col).border = border
            current_row += 1
        return current_row + 1

    row = _create_approver_section(row, 'COMPLETED BY (ASSESSOR)', '4472C4')
    row = _create_approver_section(row, 'REVIEWED BY (INFORMATION SECURITY OFFICER)', '4472C4')
    row = _create_approver_section(row, 'APPROVED BY (CISO)', '003366')

    # FINAL DECISION (GS-AS-004)
    ws[f'A{row}'] = 'FINAL DECISION:'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'A{row}'].border = border
    ws.merge_cells(f'B{row}:E{row}')
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws.cell(row=row, column=col).border = border
    decision_dv = DataValidation(
        type='list',
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(f'B{row}')

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'NEXT REVIEW DETAILS'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    row += 1
    for label in ['Next Review Date:', 'Review Responsible:', 'Special Considerations:']:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        row += 1

    # Column widths and freeze
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.freeze_panes = 'A3'

    return ws

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        ws.data_validations.dataValidation = [
            dv for dv in list(ws.data_validations.dataValidation)
            if dv.sqref
        ]


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info(f"\n{'='*70}")
    logger.info(f"GENERATING: {WORKBOOK_TITLE}")
    logger.info(f"{'='*70}")
    logger.info(f"Version: {VERSION}")
    logger.info(f"Controls: {CONTROLS}")
    logger.info(f"Assessment ID: {DOCUMENT_ID}")
    logger.info(f"{'='*70}\n")
    
    # Create workbook
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.remove(wb.active)
    
    # Create all worksheets in order
    logger.info("Creating worksheets...")
    create_instructions_sheet(wb.create_sheet())
    logger.info("  Instructions & Legend")

    create_system_inventory_sheet(wb)
    logger.info("  System Inventory")

    create_capability_assessment_sheet(wb)
    logger.info("  Capability Assessment")

    create_compliance_matrix_sheet(wb)
    logger.info("  Compliance Matrix")

    create_gap_analysis_sheet(wb)
    logger.info("  Gap Analysis")

    create_evidence_register(wb)
    logger.info("  Evidence Register")

    create_summary_dashboard_sheet(wb)
    logger.info("  Summary Dashboard")

    create_approval_sheet(wb)
    logger.info("  Approval Sign-Off")
    
    # Save workbook
    finalize_validations(wb)
    wb.save(output_path)
    # Summary
    logger.info(f"\n{'='*70}")
    logger.info("WORKBOOK GENERATED SUCCESSFULLY")
    logger.info(f"{'='*70}")
    logger.info(f"Filename: {output_path.name}")
    logger.info(f"Worksheets: {len(wb.sheetnames)}")
    logger.info("\nWorksheet Details:")
    logger.info("  • Instructions & Legend (comprehensive usage guide)")
    logger.info("  • Summary Dashboard (executive dashboard with all metrics)")
    logger.info("  • System Inventory (110 rows: 10 examples + 100 data entry)")
    logger.info("  • Capability Assessment (110 rows with automatic formulas)")
    logger.info("  • Compliance Matrix (110 rows with automatic compliance calculations)")
    logger.info("  • Gap Analysis (110 rows: 10 examples + 100 gap tracking)")
    logger.info("  • Evidence Register (1 sample row + 100 empty rows)")
    logger.info("  • Approval Sign-Off (3-level workflow: Assessor to ISO to CISO)")
    logger.info(f"\n{'='*70}")
    logger.info("{CHECK} AUDIT-READY FEATURES:")
    logger.info("  • Evidence tracking (minimum 5 items required)")
    logger.info("  • 3-level approval workflow (Assessor → ISO → CISO)")
    logger.info("  • Comprehensive data validations (10 dropdown types)")
    logger.info("  • Auto-calculated compliance metrics (RPO/RTO requirement vs capability)")
    logger.info("  • Risk-based gap prioritization (criticality × gap severity)")
    logger.info("  • Professional styling without Excel repair warnings")
    logger.info(f"{'='*70}\n")
    
def main():
    try:
        create_workbook(_wkbk_dir / OUTPUT_FILENAME)
    except Exception as e:
        logger.info(f"\n{'='*70}")
        logger.error("{XMARK} ERROR: Failed to generate workbook")
        logger.info(f"{'='*70}")
        logger.error(f"Error details: {str(e)}")
        logger.error(f"Error type: {type(e).__name__}")
        import traceback
        logger.info(f"\nFull traceback:")
        traceback.print_exc()
        logger.info(f"{'='*70}\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
