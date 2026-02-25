#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.30.S5 - BC/DR Compliance Dashboard Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.5.30 (Backup), A.8.14 (Redundancy), A.5.30 (ICT BC)
Consolidation Dashboard: Executive BC/DR Readiness Overview

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific dashboard requirements, executive reporting needs,
and assessment workbook structures.

Key customization areas:
1. Dashboard metrics and KPIs (align with executive priorities)
2. Compliance scoring methodology (weighting of different dimensions)
3. Risk severity thresholds (critical/high/medium/low definitions)
4. External workbook linking paths (match your file organization)
5. Visualization preferences (colors, conditional formatting rules)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.30/14/5.30 BC/DR Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel dashboard that consolidates data
from all four BC/DR assessment domains (Backup, Redundancy, RPO/RTO Compliance,
Testing Results) into executive-level BC/DR readiness reporting.

**Purpose:**
Provides executive oversight of BC/DR compliance status across ISO 27001:2022
Controls A.5.30, A.8.14, and A.5.30, enabling evidence-based decision-making
for BC/DR investments and risk acceptance.

**Consolidation Scope:**
- Overall BC/DR maturity scoring
- Backup coverage and compliance metrics (from Domain 1)
- Redundancy coverage and SPOF metrics (from Domain 2)
- RPO/RTO compliance and gap metrics (from Domain 3)
- Testing compliance and results metrics (from Domain 4)
- Critical gap identification and prioritization
- Trend analysis (improving vs. degrading)
- Executive action items and recommendations
- Evidence summary for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Dashboard guidance and metric definitions
2. Executive Dashboard - High-level BC/DR readiness overview
3. Backup Coverage Summary - Metrics from Domain 1 assessment
4. Redundancy & SPOF Summary - Metrics from Domain 2 assessment
5. RPO/RTO Compliance Summary - Metrics from Domain 3 assessment
6. Testing Results Summary - Metrics from Domain 4 assessment
7. Critical Gaps Register - Consolidated high-priority gaps
8. Evidence Checklist - Audit evidence availability summary
9. Approval & Sign-Off - Executive review and approval workflow

**Key Features:**
- External workbook formula linking to source assessments
- Automated metric calculations from source data
- Conditional formatting for status visualization
- Executive KPIs and trend indicators
- Critical gap highlighting and prioritization
- Protected formulas with unprotected annotation cells
- Evidence completeness tracking
- Executive approval workflow
- Professional dashboard styling

**Integration:**
This dashboard consolidates data from:
- ISMS-IMP-A.5.30.S1.xlsx (Backup Inventory Assessment)
- ISMS-IMP-A.5.30.S2.xlsx (Redundancy Analysis Assessment)
- ISMS-IMP-A.5.30.S3.xlsx (RPO/RTO Compliance Matrix)
- ISMS-IMP-A.5.30.S4.xlsx (BC/DR Testing Results Tracker)

All source workbooks must be normalized (using normalize_a530_assessments.py)
and located in the same directory as the dashboard.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl --break-system-packages

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a530_5_compliance_dashboard.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a530_5_compliance_dashboard.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a530_5_compliance_dashboard.py --date 20250125

Output:
    File: ISMS_Assessment_BCDR_Compliance_Dashboard.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Ensure all four source assessment workbooks are completed
    2. Run normalize_a530_assessments.py to normalize source file names
    3. Place dashboard and normalized sources in same directory
    4. Open dashboard in Excel
    5. Click "Enable Content" to allow external workbook links
    6. Click "Update Links" to refresh data from source workbooks
    7. Review Executive Dashboard for overall BC/DR readiness
    8. Analyze critical gaps requiring executive attention
    9. Document executive decisions on gap remediation priorities
    10. Obtain executive approval and sign-off

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.5.30, A.8.14, A.5.30
Assessment Domain:    Consolidation Dashboard (Executive Reporting)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.30-8.13-14: BC/DR Framework Policy (Governance)
    - ISMS-IMP-A.5.30-8.13-14-S5: BC/DR Assessment Guide
    - ISMS-IMP-A.5.30.S1: Backup Inventory Assessment (Source Data - Domain 1)
    - ISMS-IMP-A.5.30.S2: Redundancy Analysis Assessment (Source Data - Domain 2)
    - ISMS-IMP-A.5.30.S3: RPO/RTO Compliance Matrix (Source Data - Domain 3)
    - ISMS-IMP-A.5.30.S4: BC/DR Testing Results Tracker (Source Data - Domain 4)
    - normalize_a530_assessments.py: File normalization utility (prerequisite)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full dashboard framework per ISMS-IMP-A.5.30-8.13-14 specification
    - Consolidates data from all four BC/DR assessment domains
    - Provides executive-level BC/DR readiness reporting
    - Includes automated metric calculations and gap prioritization

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Dashboard vs. Source Data:**
This dashboard CONSOLIDATES data from source assessments - it does NOT replace them.
Source assessments (Domains 1-4) contain detailed technical data. Dashboard
provides executive summary. Both are required for comprehensive BC/DR governance.

**External Workbook Linking:**
Dashboard uses Excel external workbook formulas to link to source assessments:
- Advantage: Real-time updates when source data changes
- Requirement: Dashboard and sources must be in same directory
- Security: Excel may prompt to enable external links on first open

If external linking doesn't work (security policies, SharePoint, etc.), use
alternative consolidation approach (manual data copy or consolidation script).

**Normalization Requirement:**
Source assessments must be normalized using normalize_a530_assessments.py:
- Original files: ISMS_Assessment_Backup_Inventory_20250125.xlsx
- Normalized files: ISMS-IMP-A.5.30.S1.xlsx (no date suffix)

Dashboard links to normalized filenames. Without normalization, links will break.

**BC/DR Maturity Scoring:**
Dashboard calculates overall BC/DR maturity based on:
- Backup coverage percentage
- RPO compliance rate
- Redundancy coverage for critical systems
- Testing compliance rate
- Critical gap count

Customize weighting based on your organization's priorities.

**Critical Gap Definition:**
Critical gaps requiring executive attention:
- Tier 1 (Critical) systems without backup or redundancy
- Large RPO/RTO gaps (technical capability << business requirement)
- Failed or overdue testing for critical systems
- Identified SPOFs for critical systems without remediation plan

**Audit Considerations:**
This dashboard provides audit evidence per ISO 27001:2022 requirements:
- Demonstrates systematic BC/DR assessment
- Shows executive oversight and approval
- Documents gap remediation priorities
- Tracks evidence completeness

Auditors expect to see both dashboard AND detailed source assessments.

**Data Protection:**
Dashboard contains aggregated sensitive information including:
- Overall BC/DR readiness status
- Critical gap summary
- Executive decisions on risk acceptance

Handle in accordance with your organization's data classification policies.

**Maintenance:**
Review and update dashboard:
- Monthly: Refresh from source assessments to track progress
- Quarterly: Executive review of BC/DR readiness
- Semi-annually: Validate metric definitions still align with priorities
- Annually: Complete dashboard redesign if needed
- Ad-hoc: After major infrastructure changes or BC/DR incidents

**Quality Assurance:**
Have CISO, BC/DR Manager, and executive leadership validate dashboard metrics
before using for board reporting or audit evidence.

**Regulatory Alignment:**
Ensure dashboard metrics align with applicable regulatory reporting:
- Finance: DORA ICT risk management reporting requirements
- Healthcare: HIPAA contingency plan documentation
- Critical Infrastructure: Sector-specific resilience reporting

**Executive Communication:**
Dashboard is designed for executive audience:
- Use traffic light colors (red/yellow/green) sparingly and meaningfully
- Focus on actionable insights, not raw data
- Highlight trends (improving vs. degrading)
- Document decisions made based on dashboard data

**Limitations:**
Dashboard is only as good as source data quality:
- Garbage in = garbage out
- Validate source assessments before trusting dashboard
- Regular audits of source data accuracy
- Document assumptions and data quality issues

**Alternative to External Linking:**
If external workbook linking doesn't work in your environment:
- Option 1: Manual data consolidation (copy/paste from sources)
- Option 2: Consolidation script (programmatic data aggregation)
- Option 3: Hybrid approach (some formulas, some manual)

Document which approach you're using and why.

================================================================================
"""

# =============================================================================
# Standard Library Imports
# =============================================================================
import logging
from datetime import datetime
from pathlib import Path
import os
import sys

# =============================================================================
# Third-Party Imports
# =============================================================================
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Logging Configuration
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.30.S5"
WORKBOOK_NAME = "BC/DR Compliance Dashboard"
CONTROL_ID = "A.5.30-8.13-14"
CONTROL_NAME = "Information Backup and Redundancy"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames (sortable)

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

WORKBOOK_TITLE = "BC/DR Consolidated Dashboard"
VERSION = "1.0"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
CONTROLS = "A.5.30, A.8.14, A.5.30"
ASSESSMENT_ID = "BCDR-DASHBOARD-005"

# Expected normalized workbook filenames
REQUIRED_WORKBOOKS = {
    'backup': 'BCDR_1_Backup_Inventory.xlsx',
    'redundancy': 'BCDR_2_Redundancy_Analysis.xlsx',
    'rporto': 'BCDR_3_RPO_RTO_Compliance.xlsx',
    'testing': 'BCDR_4_Testing_Results.xlsx'
}

# Color scheme (consistent with reference implementations)
HEADER_FILL = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
COLUMN_HEADER_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
BLUE_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

# Fonts
HEADER_FONT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
SUBHEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
BOLD_FONT = Font(name='Calibri', size=10, bold=True)
NORMAL_FONT = Font(name='Calibri', size=10)

# Borders
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================



# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠️  Warning sign
CHART = '\U0001F4CA' # 📊 Chart
TARGET = '\U0001F3AF' # 🎯 Target
SHIELD = '\U0001F6E1' # 🛡️  Shield
LOCK = '\U0001F512'   # 🔒 Lock
DISK = '\U0001F4BE'   # 💾 Floppy Disk
GLOBE = '\U0001F310'  # 🌐 Globe
SEARCH = '\U0001F50D' # 🔍 Magnifying Glass
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

def apply_style(cell, font=None, fill=None, alignment=None, border=None):
    """Apply multiple styles to a cell"""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border

def safe_read_cell(wb, sheet_name, cell_ref, default=0):
    """Safely read a cell value from workbook, return default if error"""
    try:
        if sheet_name not in wb.sheetnames:
            return default
        ws = wb[sheet_name]
        value = ws[cell_ref].value
        if value is None:
            return default
        # Handle percentage values
        if isinstance(value, str) and '%' in value:
            return float(value.strip('%')) / 100
        return value
    except Exception as e:  # Return default on parse errors
        return default

def create_base_validations(ws):
    """Create base data validation objects for dropdowns"""
    validations = {
        'maturity_level': DataValidation(
            type="list",
            formula1='"Level 1 - Initial,Level 2 - Managed,Level 3 - Defined,Level 4 - Optimized"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1='"Draft,Under Review,Final,Requires Remediation"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected,Requires Rework"',
            allow_blank=False
        ),
        'recommendation': DataValidation(
            type="list",
            formula1='"Approve,Approve with Conditions,Reject,Require Rework"',
            allow_blank=False
        ),
    }
    
    # Add all validations to worksheet
    for dv in validations.values():
        ws.add_data_validation(dv)
    
    return validations

# ============================================================================
# DATA EXTRACTION FUNCTIONS
# ============================================================================

def load_assessment_data():
    """
    Load and extract data from all 4 normalized assessment workbooks.
    
    Returns:
        dict: Consolidated metrics from all assessments
    """
    data = {
        'backup': {},
        'redundancy': {},
        'rporto': {},
        'testing': {},
        'missing': []
    }
    
    # Load Backup Inventory (WB1)
    try:
        wb1 = load_workbook(REQUIRED_WORKBOOKS['backup'], read_only=True, data_only=True)
        data['backup'] = {
            'total_systems': safe_read_cell(wb1, 'Summary', 'B5', 0),
            'backed_up': safe_read_cell(wb1, 'Summary', 'B6', 0),
            'coverage': safe_read_cell(wb1, 'Summary', 'B7', 0),
            'critical_systems': safe_read_cell(wb1, 'Summary', 'B8', 0),
            'critical_backed_up': safe_read_cell(wb1, 'Summary', 'B9', 0),
            'critical_coverage': safe_read_cell(wb1, 'Summary', 'B10', 0),
            'rpo_compliant': safe_read_cell(wb1, 'Summary', 'B13', 0),
            'rpo_compliance_rate': safe_read_cell(wb1, 'Summary', 'B15', 0),
            '321_full_compliance': safe_read_cell(wb1, 'Summary', 'B22', 0),
            '321_compliance_rate': safe_read_cell(wb1, 'Summary', 'B27', 0),
            'evidence_count': safe_read_cell(wb1, 'Summary', 'B28', 0),
        }
        wb1.close()
    except FileNotFoundError:
        data['missing'].append(REQUIRED_WORKBOOKS['backup'])
    except Exception as e:
        logger.error(f"    ⚠️  Error loading {REQUIRED_WORKBOOKS['backup']}: {e}")
        data['missing'].append(REQUIRED_WORKBOOKS['backup'])
    
    # Load Redundancy Analysis (WB2)
    try:
        wb2 = load_workbook(REQUIRED_WORKBOOKS['redundancy'], read_only=True, data_only=True)
        data['redundancy'] = {
            'total_critical_systems': safe_read_cell(wb2, 'Summary', 'B5', 0),
            'with_redundancy': safe_read_cell(wb2, 'Summary', 'B6', 0),
            'redundancy_coverage': safe_read_cell(wb2, 'Summary', 'B7', 0),
            'rto_compliant': safe_read_cell(wb2, 'Summary', 'B10', 0),
            'rto_compliance_rate': safe_read_cell(wb2, 'Summary', 'B12', 0),
            'spof_identified': safe_read_cell(wb2, 'Summary', 'B15', 0),
            'spof_mitigated': safe_read_cell(wb2, 'Summary', 'B16', 0),
            'spof_open': safe_read_cell(wb2, 'Summary', 'B17', 0),
            'failover_tested': safe_read_cell(wb2, 'Summary', 'B22', 0),
            'failover_success_rate': safe_read_cell(wb2, 'Summary', 'B25', 0),
            'evidence_count': safe_read_cell(wb2, 'Summary', 'B28', 0),
        }
        wb2.close()
    except FileNotFoundError:
        data['missing'].append(REQUIRED_WORKBOOKS['redundancy'])
    except Exception as e:
        logger.error(f"    ⚠️  Error loading {REQUIRED_WORKBOOKS['redundancy']}: {e}")
        data['missing'].append(REQUIRED_WORKBOOKS['redundancy'])
    
    # Load RPO/RTO Compliance (WB3)
    try:
        wb3 = load_workbook(REQUIRED_WORKBOOKS['rporto'], read_only=True, data_only=True)
        data['rporto'] = {
            'total_systems': safe_read_cell(wb3, 'Summary', 'B5', 0),
            'full_compliance': safe_read_cell(wb3, 'Summary', 'B6', 0),
            'partial_compliance': safe_read_cell(wb3, 'Summary', 'B7', 0),
            'non_compliant': safe_read_cell(wb3, 'Summary', 'B8', 0),
            'unknown': safe_read_cell(wb3, 'Summary', 'B9', 0),
            'overall_compliance_rate': safe_read_cell(wb3, 'Summary', 'B10', 0),
            'tier1_systems': safe_read_cell(wb3, 'Summary', 'B13', 0),
            'tier1_compliant': safe_read_cell(wb3, 'Summary', 'B14', 0),
            'tier1_compliance_rate': safe_read_cell(wb3, 'Summary', 'B15', 0),
            'total_gaps': safe_read_cell(wb3, 'Summary', 'B22', 0),
            'critical_gaps': safe_read_cell(wb3, 'Summary', 'B23', 0),
            'open_gaps': safe_read_cell(wb3, 'Summary', 'B29', 0),
            'evidence_count': safe_read_cell(wb3, 'Summary', 'B40', 0),
        }
        wb3.close()
    except FileNotFoundError:
        data['missing'].append(REQUIRED_WORKBOOKS['rporto'])
    except Exception as e:
        logger.error(f"    ⚠️  Error loading {REQUIRED_WORKBOOKS['rporto']}: {e}")
        data['missing'].append(REQUIRED_WORKBOOKS['rporto'])
    
    # Load Testing Results (WB4)
    try:
        wb4 = load_workbook(REQUIRED_WORKBOOKS['testing'], read_only=True, data_only=True)
        data['testing'] = {
            'total_scheduled': safe_read_cell(wb4, 'Summary', 'B5', 0),
            'completed_success': safe_read_cell(wb4, 'Summary', 'B6', 0),
            'partial_success': safe_read_cell(wb4, 'Summary', 'B7', 0),
            'failed': safe_read_cell(wb4, 'Summary', 'B8', 0),
            'success_rate': safe_read_cell(wb4, 'Summary', 'B12', 0),
            'tier1_systems': safe_read_cell(wb4, 'Summary', 'B15', 0),
            'tier1_compliant': safe_read_cell(wb4, 'Summary', 'B16', 0),
            'tier1_compliance_rate': safe_read_cell(wb4, 'Summary', 'B17', 0),
            'total_issues': safe_read_cell(wb4, 'Summary', 'B24', 0),
            'critical_issues': safe_read_cell(wb4, 'Summary', 'B25', 0),
            'open_issues': safe_read_cell(wb4, 'Summary', 'B29', 0),
            'evidence_count': safe_read_cell(wb4, 'Summary', 'B40', 0),
        }
        wb4.close()
    except FileNotFoundError:
        data['missing'].append(REQUIRED_WORKBOOKS['testing'])
    except Exception as e:
        logger.error(f"    ⚠️  Error loading {REQUIRED_WORKBOOKS['testing']}: {e}")
        data['missing'].append(REQUIRED_WORKBOOKS['testing'])
    
    return data

def calculate_maturity_score(data):
    """
    Calculate overall BC/DR maturity score (0-100%).
    
    Maturity dimensions (25 points each):
    1. Backup Coverage & Compliance (25%)
    2. Redundancy Coverage & Compliance (25%)
    3. RPO/RTO Alignment (25%)
    4. Testing Compliance (25%)
    
    Returns:
        tuple: (score, level_name, level_description)
    """
    scores = {
        'backup': 0,
        'redundancy': 0,
        'rporto': 0,
        'testing': 0
    }
    
    # 1. Backup Coverage & Compliance (25 points)
    if data['backup']:
        backup_score = 0
        # Critical system coverage (10 points)
        backup_score += data['backup'].get('critical_coverage', 0) * 10
        # Overall coverage (5 points)
        backup_score += data['backup'].get('coverage', 0) * 5
        # RPO compliance (5 points)
        backup_score += data['backup'].get('rpo_compliance_rate', 0) * 5
        # 3-2-1-1-0 compliance (5 points)
        backup_score += data['backup'].get('321_compliance_rate', 0) * 5
        scores['backup'] = min(backup_score, 25)
    
    # 2. Redundancy Coverage & Compliance (25 points)
    if data['redundancy']:
        redundancy_score = 0
        # Redundancy coverage (10 points)
        redundancy_score += data['redundancy'].get('redundancy_coverage', 0) * 10
        # RTO compliance (10 points)
        redundancy_score += data['redundancy'].get('rto_compliance_rate', 0) * 10
        # Failover testing success (5 points)
        redundancy_score += data['redundancy'].get('failover_success_rate', 0) * 5
        scores['redundancy'] = min(redundancy_score, 25)
    
    # 3. RPO/RTO Alignment (25 points)
    if data['rporto']:
        rporto_score = 0
        # Overall RPO/RTO compliance (15 points)
        rporto_score += data['rporto'].get('overall_compliance_rate', 0) * 15
        # Tier 1 compliance (10 points)
        rporto_score += data['rporto'].get('tier1_compliance_rate', 0) * 10
        scores['rporto'] = min(rporto_score, 25)
    
    # 4. Testing Compliance (25 points)
    if data['testing']:
        testing_score = 0
        # Test success rate (15 points)
        testing_score += data['testing'].get('success_rate', 0) * 15
        # Tier 1 testing compliance (10 points)
        testing_score += data['testing'].get('tier1_compliance_rate', 0) * 10
        scores['testing'] = min(testing_score, 25)
    
    # Total score (0-100)
    total_score = sum(scores.values())
    
    # Determine maturity level
    if total_score >= 85:
        level = 4
        level_name = "Level 4 - Optimized"
        level_desc = "BC/DR capabilities are continuously optimized and meet all requirements"
    elif total_score >= 70:
        level = 3
        level_name = "Level 3 - Defined"
        level_desc = "BC/DR processes are defined and consistently followed"
    elif total_score >= 50:
        level = 2
        level_name = "Level 2 - Managed"
        level_desc = "BC/DR processes exist but with gaps in coverage or compliance"
    else:
        level = 1
        level_name = "Level 1 - Initial"
        level_desc = "BC/DR capabilities are ad-hoc with significant gaps"
    
    return (total_score, level, level_name, level_desc, scores)

# ============================================================================
# WORKSHEET: INSTRUCTIONS
# ============================================================================

def create_instructions_sheet(wb):
    """Create comprehensive instructions worksheet"""
    ws = wb.create_sheet(title="Instructions", index=0)
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f'ISMS BC/DR Assessment - {WORKBOOK_TITLE}'
    apply_style(ws['A1'], font=Font(name='Calibri', size=16, bold=True, color='003366'),
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Version {VERSION} | Controls: {CONTROLS} | Assessment ID: {ASSESSMENT_ID}'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(italic=True, size=11)
    
    # Purpose
    row = 4
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'PURPOSE'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    purpose_text = [
        'This dashboard consolidates metrics from all 4 BC/DR assessment workbooks into a single executive view.',
        'It provides an overall BC/DR maturity score (0-100%) and identifies critical gaps requiring attention.',
        '',
        'PREREQUISITE: All 4 normalized assessment workbooks must be present in the same directory.',
    ]
    
    for text in purpose_text:
        ws[f'A{row}'] = text
        ws[f'A{row}'].font = NORMAL_FONT
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Source Workbooks
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'SOURCE WORKBOOKS (REQUIRED)'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    source_workbooks = [
        ('BCDR_1_Backup_Inventory.xlsx', 'Backup coverage, RPO compliance, 3-2-1-1-0 rule'),
        ('BCDR_2_Redundancy_Analysis.xlsx', 'Redundancy coverage, RTO compliance, SPOF analysis'),
        ('BCDR_3_RPO_RTO_Compliance.xlsx', 'RPO/RTO alignment, gap analysis'),
        ('BCDR_4_Testing_Results.xlsx', 'Testing schedule, results, issues'),
    ]
    
    for filename, description in source_workbooks:
        ws[f'A{row}'] = filename
        ws[f'B{row}'] = description
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'B{row}'].font = NORMAL_FONT
        row += 1
    
    # Worksheets Description
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'WORKSHEET DESCRIPTIONS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    worksheet_descriptions = [
        ('Instructions:', 'This sheet - comprehensive usage guide'),
        ('Executive_Dashboard:', 'Overall BC/DR maturity score and key metrics'),
        ('Detailed_Metrics:', 'Detailed metrics from all 4 assessments'),
        ('Gap_Summary:', 'Consolidated critical gaps from all assessments'),
        ('Evidence_Checklist:', 'Evidence completeness tracking'),
        ('Approval_Sign_Off:', '3-level approval workflow (Assessor → ISO → CISO)'),
    ]
    
    for label, description in worksheet_descriptions:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = description
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'B{row}'].font = NORMAL_FONT
        row += 1
    
    # Maturity Scoring
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'MATURITY SCORING METHODOLOGY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    maturity_scoring = [
        'Overall BC/DR Maturity Score: 0-100% (calculated from 4 dimensions)',
        '',
        'DIMENSION 1: Backup Coverage & Compliance (25 points)',
        '  • Critical system backup coverage (10 points)',
        '  • Overall backup coverage (5 points)',
        '  • RPO compliance rate (5 points)',
        '  • 3-2-1-1-0 compliance rate (5 points)',
        '',
        'DIMENSION 2: Redundancy Coverage & Compliance (25 points)',
        '  • Redundancy coverage for critical systems (10 points)',
        '  • RTO compliance rate (10 points)',
        '  • Failover testing success rate (5 points)',
        '',
        'DIMENSION 3: RPO/RTO Alignment (25 points)',
        '  • Overall RPO/RTO compliance rate (15 points)',
        '  • Tier 1 system compliance rate (10 points)',
        '',
        'DIMENSION 4: Testing Compliance (25 points)',
        '  • Test success rate (15 points)',
        '  • Tier 1 testing compliance (10 points)',
    ]
    
    for text in maturity_scoring:
        ws[f'A{row}'] = text
        if f'{BULLET}' in text or 'DIMENSION' in text:
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold='DIMENSION' in text)
        else:
            ws[f'A{row}'].font = NORMAL_FONT
        row += 1
    
    # Maturity Levels
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'MATURITY LEVELS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    maturity_levels = [
        ('Level 4 - Optimized (85-100%):', 'BC/DR capabilities continuously optimized, meet all requirements'),
        ('Level 3 - Defined (70-84%):', 'BC/DR processes defined and consistently followed'),
        ('Level 2 - Managed (50-69%):', 'BC/DR processes exist but with coverage/compliance gaps'),
        ('Level 1 - Initial (0-49%):', 'BC/DR capabilities ad-hoc with significant gaps'),
    ]
    
    for level, description in maturity_levels:
        ws[f'A{row}'] = level
        ws[f'B{row}'] = description
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'B{row}'].font = NORMAL_FONT
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 60
    for col in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15
    
    return ws

# ============================================================================
# WORKSHEET: EXECUTIVE DASHBOARD
# ============================================================================

def create_executive_dashboard_sheet(wb, data, maturity_results):
    """Create Executive Dashboard worksheet with consolidated metrics"""
    ws = wb.create_sheet(title="Executive_Dashboard", index=1)
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = 'BC/DR MATURITY DASHBOARD'
    apply_style(ws['A1'], font=Font(name='Calibri', size=18, bold=True, color='003366'),
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Assessment Date: {datetime.now().strftime("%d.%m.%Y")} | Overall BC/DR Maturity Score: {maturity_results[0]:.1f}%'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(bold=True, size=12)
    
    # Maturity Score Section
    row = 4
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'OVERALL BC/DR MATURITY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = 'Maturity Score:'
    apply_style(ws[f'A{row}'], font=Font(name='Calibri', size=14, bold=True))
    
    ws.merge_cells(f'C{row}:D{row}')
    ws[f'C{row}'] = f'{maturity_results[0]:.1f}%'
    apply_style(ws[f'C{row}'], font=Font(name='Calibri', size=20, bold=True, color='003366'))
    ws[f'C{row}'].alignment = Alignment(horizontal='center')
    
    # Color code based on score
    if maturity_results[0] >= 85:
        fill_color = GREEN_FILL
    elif maturity_results[0] >= 70:
        fill_color = BLUE_FILL
    elif maturity_results[0] >= 50:
        fill_color = YELLOW_FILL
    else:
        fill_color = RED_FILL
    ws[f'C{row}'].fill = fill_color
    ws[f'D{row}'].fill = fill_color
    
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = 'Maturity Level:'
    apply_style(ws[f'A{row}'], font=BOLD_FONT)
    
    ws.merge_cells(f'C{row}:F{row}')
    ws[f'C{row}'] = maturity_results[2]
    apply_style(ws[f'C{row}'], font=Font(name='Calibri', size=12, bold=True))
    
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = maturity_results[3]
    ws[f'A{row}'].font = Font(italic=True, size=10)
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    
    # Dimension Scores
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'MATURITY DIMENSION SCORES'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    dimension_scores = [
        ('Backup Coverage & Compliance', maturity_results[4]['backup'], 25),
        ('Redundancy Coverage & Compliance', maturity_results[4]['redundancy'], 25),
        ('RPO/RTO Alignment', maturity_results[4]['rporto'], 25),
        ('Testing Compliance', maturity_results[4]['testing'], 25),
    ]
    
    for dimension, score, max_score in dimension_scores:
        ws[f'A{row}'] = dimension
        ws[f'B{row}'] = f'{score:.1f}/{max_score}'
        ws[f'C{row}'] = f'{(score/max_score*100):.0f}%'
        
        apply_style(ws[f'A{row}'], font=BOLD_FONT, border=THIN_BORDER)
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=11), border=THIN_BORDER)
        apply_style(ws[f'C{row}'], font=Font(bold=True, size=11), border=THIN_BORDER)
        
        # Color code percentage
        pct = score/max_score
        if pct >= 0.85:
            ws[f'C{row}'].fill = GREEN_FILL
        elif pct >= 0.70:
            ws[f'C{row}'].fill = BLUE_FILL
        elif pct >= 0.50:
            ws[f'C{row}'].fill = YELLOW_FILL
        else:
            ws[f'C{row}'].fill = RED_FILL
        
        row += 1
    
    # Key Metrics Summary
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'KEY METRICS SUMMARY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    
    # Backup metrics
    if data['backup']:
        ws[f'A{row}'] = 'Backup Coverage (Critical Systems):'
        ws[f'B{row}'] = f"{data['backup'].get('critical_coverage', 0):.1%}"
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1
    
    # Redundancy metrics
    if data['redundancy']:
        ws[f'A{row}'] = 'Redundancy Coverage (Critical Systems):'
        ws[f'B{row}'] = f"{data['redundancy'].get('redundancy_coverage', 0):.1%}"
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1
    
    # RPO/RTO metrics
    if data['rporto']:
        ws[f'A{row}'] = 'RPO/RTO Compliance (Tier 1):'
        ws[f'B{row}'] = f"{data['rporto'].get('tier1_compliance_rate', 0):.1%}"
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1
    
    # Testing metrics
    if data['testing']:
        ws[f'A{row}'] = 'Test Success Rate:'
        ws[f'B{row}'] = f"{data['testing'].get('success_rate', 0):.1%}"
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1
    
    # Critical Gaps Summary
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'CRITICAL GAPS & ISSUES'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    
    total_critical_gaps = 0
    total_open_gaps = 0
    total_critical_issues = 0
    total_open_issues = 0
    
    if data['rporto']:
        total_critical_gaps += data['rporto'].get('critical_gaps', 0)
        total_open_gaps += data['rporto'].get('open_gaps', 0)
    
    if data['testing']:
        total_critical_issues += data['testing'].get('critical_issues', 0)
        total_open_issues += data['testing'].get('open_issues', 0)
    
    critical_metrics = [
        ('🔴 Critical Gaps (RPO/RTO):', total_critical_gaps),
        ('🔴 Open Gaps Requiring Remediation:', total_open_gaps),
        ('🔴 Critical Issues (Testing):', total_critical_issues),
        ('🔴 Open Issues:', total_open_issues),
    ]
    
    for label, value in critical_metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        apply_style(ws[f'A{row}'], font=Font(name='Calibri', size=10, bold=True, color='C00000'))
        apply_style(ws[f'B{row}'], font=Font(name='Calibri', size=11, bold=True, color='C00000'))
        if value > 0:
            ws[f'B{row}'].fill = RED_FILL
        row += 1
    
    # Next Steps
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'PRIORITY ACTIONS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    priority_actions = [
        '1. Review Gap_Summary worksheet for all critical gaps',
        '2. Address RPO/RTO compliance gaps for Tier 1 systems',
        '3. Remediate all critical issues discovered during testing',
        '4. Complete overdue testing for critical systems',
        '5. Review Detailed_Metrics worksheet for specific metrics',
        '6. Ensure Evidence_Checklist is complete (minimum 5 items per control)',
        '7. Complete Approval_Sign_Off workflow (3 levels)',
    ]
    
    for action in priority_actions:
        ws[f'A{row}'] = action
        ws[f'A{row}'].font = NORMAL_FONT
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    for col in ['D', 'E', 'F']:
        ws.column_dimensions[col].width = 15
    
    return ws

# ============================================================================
# WORKSHEET: DETAILED METRICS
# ============================================================================

def create_detailed_metrics_sheet(wb, data):
    """Create Detailed Metrics worksheet with all metrics from 4 workbooks"""
    ws = wb.create_sheet(title="Detailed_Metrics")
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = 'DETAILED BC/DR METRICS'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Backup Metrics
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'BACKUP COVERAGE & COMPLIANCE (A.5.30)'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    if data['backup']:
        backup_metrics = [
            ('Total Systems Assessed', data['backup'].get('total_systems', 0)),
            ('Systems with Backup', data['backup'].get('backed_up', 0)),
            ('Overall Backup Coverage', f"{data['backup'].get('coverage', 0):.1%}"),
            ('Critical Systems (Tier 1)', data['backup'].get('critical_systems', 0)),
            ('Critical Systems Backed Up', data['backup'].get('critical_backed_up', 0)),
            ('Critical System Coverage', f"{data['backup'].get('critical_coverage', 0):.1%}"),
            ('RPO Compliant Systems', data['backup'].get('rpo_compliant', 0)),
            ('RPO Compliance Rate', f"{data['backup'].get('rpo_compliance_rate', 0):.1%}"),
            ('3-2-1-1-0 Full Compliance', data['backup'].get('321_full_compliance', 0)),
            ('3-2-1-1-0 Compliance Rate', f"{data['backup'].get('321_compliance_rate', 0):.1%}"),
            ('Evidence Items Collected', data['backup'].get('evidence_count', 0)),
        ]
        
        for metric, value in backup_metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            apply_style(ws[f'A{row}'], font=BOLD_FONT, border=THIN_BORDER)
            apply_style(ws[f'B{row}'], font=NORMAL_FONT, border=THIN_BORDER)
            row += 1
    else:
        ws[f'A{row}'] = f'{XMARK} BCDR_1_Backup_Inventory.xlsx not found'
        ws[f'A{row}'].font = Font(color='C00000', bold=True)
        row += 1
    
    # Redundancy Metrics
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'REDUNDANCY COVERAGE & COMPLIANCE (A.8.14)'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    if data['redundancy']:
        redundancy_metrics = [
            ('Total Critical Systems', data['redundancy'].get('total_critical_systems', 0)),
            ('Systems with Redundancy', data['redundancy'].get('with_redundancy', 0)),
            ('Redundancy Coverage', f"{data['redundancy'].get('redundancy_coverage', 0):.1%}"),
            ('RTO Compliant Systems', data['redundancy'].get('rto_compliant', 0)),
            ('RTO Compliance Rate', f"{data['redundancy'].get('rto_compliance_rate', 0):.1%}"),
            ('SPOFs Identified', data['redundancy'].get('spof_identified', 0)),
            ('SPOFs Mitigated', data['redundancy'].get('spof_mitigated', 0)),
            ('SPOFs Still Open', data['redundancy'].get('spof_open', 0)),
            ('Failover Tests Completed', data['redundancy'].get('failover_tested', 0)),
            ('Failover Success Rate', f"{data['redundancy'].get('failover_success_rate', 0):.1%}"),
            ('Evidence Items Collected', data['redundancy'].get('evidence_count', 0)),
        ]
        
        for metric, value in redundancy_metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            apply_style(ws[f'A{row}'], font=BOLD_FONT, border=THIN_BORDER)
            apply_style(ws[f'B{row}'], font=NORMAL_FONT, border=THIN_BORDER)
            row += 1
    else:
        ws[f'A{row}'] = f'{XMARK} BCDR_2_Redundancy_Analysis.xlsx not found'
        ws[f'A{row}'].font = Font(color='C00000', bold=True)
        row += 1
    
    # RPO/RTO Compliance Metrics
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'RPO/RTO COMPLIANCE (A.5.30, A.8.14, A.5.30)'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    if data['rporto']:
        rporto_metrics = [
            ('Total Systems Assessed', data['rporto'].get('total_systems', 0)),
            ('Full Compliance (RPO & RTO)', data['rporto'].get('full_compliance', 0)),
            ('Partial Compliance', data['rporto'].get('partial_compliance', 0)),
            ('Non-Compliant Systems', data['rporto'].get('non_compliant', 0)),
            ('Unknown Status (Not Tested)', data['rporto'].get('unknown', 0)),
            ('Overall Compliance Rate', f"{data['rporto'].get('overall_compliance_rate', 0):.1%}"),
            ('Tier 1 Systems', data['rporto'].get('tier1_systems', 0)),
            ('Tier 1 Compliant', data['rporto'].get('tier1_compliant', 0)),
            ('Tier 1 Compliance Rate', f"{data['rporto'].get('tier1_compliance_rate', 0):.1%}"),
            ('Total Gaps Identified', data['rporto'].get('total_gaps', 0)),
            ('Critical Priority Gaps', data['rporto'].get('critical_gaps', 0)),
            ('Open Gaps', data['rporto'].get('open_gaps', 0)),
            ('Evidence Items Collected', data['rporto'].get('evidence_count', 0)),
        ]
        
        for metric, value in rporto_metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            apply_style(ws[f'A{row}'], font=BOLD_FONT, border=THIN_BORDER)
            apply_style(ws[f'B{row}'], font=NORMAL_FONT, border=THIN_BORDER)
            row += 1
    else:
        ws[f'A{row}'] = f'{XMARK} BCDR_3_RPO_RTO_Compliance.xlsx not found'
        ws[f'A{row}'].font = Font(color='C00000', bold=True)
        row += 1
    
    # Testing Metrics
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'TESTING COMPLIANCE (A.5.30, A.8.14, A.5.30)'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    if data['testing']:
        testing_metrics = [
            ('Total Tests Scheduled', data['testing'].get('total_scheduled', 0)),
            ('Tests Completed (Success)', data['testing'].get('completed_success', 0)),
            ('Tests Partial Success', data['testing'].get('partial_success', 0)),
            ('Tests Failed', data['testing'].get('failed', 0)),
            ('Test Success Rate', f"{data['testing'].get('success_rate', 0):.1%}"),
            ('Tier 1 Systems', data['testing'].get('tier1_systems', 0)),
            ('Tier 1 Testing Compliant', data['testing'].get('tier1_compliant', 0)),
            ('Tier 1 Compliance Rate', f"{data['testing'].get('tier1_compliance_rate', 0):.1%}"),
            ('Total Issues Identified', data['testing'].get('total_issues', 0)),
            ('Critical Issues', data['testing'].get('critical_issues', 0)),
            ('Open Issues', data['testing'].get('open_issues', 0)),
            ('Evidence Items Collected', data['testing'].get('evidence_count', 0)),
        ]
        
        for metric, value in testing_metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            apply_style(ws[f'A{row}'], font=BOLD_FONT, border=THIN_BORDER)
            apply_style(ws[f'B{row}'], font=NORMAL_FONT, border=THIN_BORDER)
            row += 1
    else:
        ws[f'A{row}'] = f'{XMARK} BCDR_4_Testing_Results.xlsx not found'
        ws[f'A{row}'].font = Font(color='C00000', bold=True)
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    return ws

# ============================================================================
# WORKSHEET: GAP SUMMARY
# ============================================================================

def create_gap_summary_sheet(wb, data):
    """Create Gap Summary worksheet consolidating critical gaps"""
    ws = wb.create_sheet(title="Gap_Summary")
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = 'CRITICAL GAPS & ISSUES SUMMARY'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:F2')
    ws['A2'] = 'Consolidated view of all critical gaps and issues requiring immediate attention'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'Source Assessment',
        'Gap/Issue Type',
        'System/Area',
        'Priority',
        'Description',
        'Status'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    row += 1
    
    # Instructions for user to populate
    instruction_rows = [
        ['RPO/RTO Compliance', 'RPO Gap / RTO Gap', '[System Name]', '🔴 Critical', 
         'Review BCDR_3_RPO_RTO_Compliance.xlsx → Gap_Analysis sheet', '🔴 Open'],
        ['Testing Results', 'Test Failure / Critical Issue', '[System Name]', '🔴 Critical', 
         'Review BCDR_4_Testing_Results.xlsx → Issue_Remediation sheet', '🔴 Open'],
        ['Backup Coverage', 'Missing Backup / RPO Non-Compliance', '[System Name]', '🟡 High', 
         'Review BCDR_1_Backup_Inventory.xlsx → Backup_Inventory sheet', '⏳ In Progress'],
        ['Redundancy', 'SPOF / RTO Non-Compliance', '[System Name]', '🟡 High', 
         'Review BCDR_2_Redundancy_Analysis.xlsx → SPOF_Register sheet', '🔴 Open'],
        ['', '', '', '', 
         'NOTE: Populate this sheet with actual critical gaps from the 4 assessment workbooks', ''],
    ]
    
    for row_data in instruction_rows:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_style(cell, border=THIN_BORDER)
            if 'NOTE:' in str(value):
                cell.font = Font(italic=True, bold=True, color='666666')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        row += 1
    
    # Gap Statistics
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'GAP STATISTICS (FROM SOURCE WORKBOOKS)'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    
    if data['rporto']:
        ws[f'A{row}'] = 'RPO/RTO Gaps (Critical Priority):'
        ws[f'B{row}'] = data['rporto'].get('critical_gaps', 0)
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1
        
        ws[f'A{row}'] = 'RPO/RTO Open Gaps:'
        ws[f'B{row}'] = data['rporto'].get('open_gaps', 0)
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1
    
    if data['testing']:
        ws[f'A{row}'] = 'Testing Issues (Critical Severity):'
        ws[f'B{row}'] = data['testing'].get('critical_issues', 0)
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1
        
        ws[f'A{row}'] = 'Testing Open Issues:'
        ws[f'B{row}'] = data['testing'].get('open_issues', 0)
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1
    
    if data['redundancy']:
        ws[f'A{row}'] = 'SPOFs Still Open:'
        ws[f'B{row}'] = data['redundancy'].get('spof_open', 0)
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1
    
    # Column widths
    widths = [25, 25, 20, 15, 45, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'A5'
    
    return ws

# ============================================================================
# WORKSHEET: EVIDENCE CHECKLIST
# ============================================================================

def create_evidence_checklist_sheet(wb, data):
    """Create Evidence Checklist worksheet tracking evidence from all workbooks"""
    ws = wb.create_sheet(title="Evidence_Checklist")
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = 'EVIDENCE COMPLETENESS CHECKLIST'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Track evidence completeness across all 4 BC/DR assessment workbooks'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'Control',
        'Assessment Workbook',
        'Evidence Count',
        'Min Required',
        'Status'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    row += 1
    
    # Evidence summary from each workbook
    evidence_data = [
        ('A.5.30', 'BCDR_1_Backup_Inventory.xlsx', 
         data['backup'].get('evidence_count', 0) if data['backup'] else 0, 5),
        ('A.8.14', 'BCDR_2_Redundancy_Analysis.xlsx', 
         data['redundancy'].get('evidence_count', 0) if data['redundancy'] else 0, 5),
        ('A.5.30, A.8.14, A.5.30', 'BCDR_3_RPO_RTO_Compliance.xlsx', 
         data['rporto'].get('evidence_count', 0) if data['rporto'] else 0, 5),
        ('A.5.30, A.8.14, A.5.30', 'BCDR_4_Testing_Results.xlsx', 
         data['testing'].get('evidence_count', 0) if data['testing'] else 0, 5),
    ]
    
    total_evidence = 0
    total_required = 0
    
    for control, workbook, count, required in evidence_data:
        ws[f'A{row}'] = control
        ws[f'B{row}'] = workbook
        ws[f'C{row}'] = count
        ws[f'D{row}'] = required
        
        # Status
        if count >= required:
            ws[f'E{row}'] = f'{CHECK} Sufficient'
            fill_color = GREEN_FILL
        elif count > 0:
            ws[f'E{row}'] = f'{WARNING} Partial'
            fill_color = YELLOW_FILL
        else:
            ws[f'E{row}'] = f'{XMARK} Insufficient'
            fill_color = RED_FILL
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            apply_style(ws[f'{col}{row}'], border=THIN_BORDER)
        ws[f'E{row}'].fill = fill_color
        
        total_evidence += count
        total_required += required
        row += 1
    
    # Total row
    ws[f'A{row}'] = 'TOTAL'
    ws[f'B{row}'] = 'All Assessments'
    ws[f'C{row}'] = total_evidence
    ws[f'D{row}'] = total_required
    
    if total_evidence >= total_required:
        ws[f'E{row}'] = f'{CHECK} Compliant'
        fill_color = GREEN_FILL
    else:
        ws[f'E{row}'] = f'{XMARK} Non-Compliant'
        fill_color = RED_FILL
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], font=BOLD_FONT, border=THIN_BORDER)
    ws[f'E{row}'].fill = fill_color
    
    row += 2
    
    # Evidence requirements note
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'NOTE: Each assessment workbook must have minimum 5 evidence items in its Evidence_Register'
    ws[f'A{row}'].font = Font(italic=True, size=10, color='666666')
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    
    return ws

# ============================================================================
# WORKSHEET: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_signoff(wb):
    """Create Approval_Sign_Off worksheet with 3-level approval workflow"""
    ws = wb.create_sheet(title="Approval_Sign_Off")
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = 'DASHBOARD APPROVAL & SIGN-OFF'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 30
    
    # Assessment Summary
    row = 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'ASSESSMENT SUMMARY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    summary_items = [
        ('Assessment Document:', WORKBOOK_TITLE),
        ('Assessment ID:', ASSESSMENT_ID),
        ('ISO 27001:2022 Controls:', CONTROLS),
        ('Assessment Period:', '[USER INPUT - e.g., Annual 2024]'),
        ('BC/DR Maturity Score:', '=Executive_Dashboard!C5'),
        ('Maturity Level:', '=Executive_Dashboard!C6'),
        ('Critical Gaps (RPO/RTO):', '=Executive_Dashboard!B17'),
        ('Critical Issues (Testing):', '=Executive_Dashboard!B19'),
        ('Evidence Compliance:', '=Evidence_Checklist!E9'),
        ('Assessment Status:', '[SELECT FROM DROPDOWN]'),
    ]
    
    for label, value in summary_items:
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        cell.value = value
        
        if 'USER INPUT' in str(value) or 'SELECT' in str(value):
            apply_style(cell, fill=INPUT_FILL)
        if 'DROPDOWN' in str(value):
            validations['assessment_status'].add(cell)
        
        row += 1
    
    # Assessment Completed By (Level 1)
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'LEVEL 1: ASSESSMENT COMPLETED BY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    completion_fields = [
        ('Name:', None),
        ('Role/Title:', None),
        ('Department:', None),
        ('Email:', None),
        ('Date Completed:', 'auto_date'),
        ('Signature/Initials:', None),
    ]
    
    for field, special in completion_fields:
        ws[f'A{row}'] = field
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        apply_style(cell, fill=INPUT_FILL)
        
        if special == 'auto_date':
            cell.value = '=TODAY()'
            cell.number_format = 'DD.MM.YYYY'
            cell.font = Font(italic=True, color='808080')
        
        row += 1
    
    # Reviewed By (Level 2 - ISO)
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'LEVEL 2: REVIEWED BY (INFORMATION SECURITY OFFICER)'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    review_fields = [
        ('Name:', None),
        ('Review Date:', 'date'),
        ('Review Notes:', 3),  # Multi-line (3 rows)
        ('Recommendation:', 'recommendation_dropdown'),
    ]
    
    for field, special in review_fields:
        ws[f'A{row}'] = field
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        apply_style(cell, fill=INPUT_FILL)
        
        if special == 'recommendation_dropdown':
            validations['recommendation'].add(cell)
        elif special == 'date':
            cell.number_format = 'DD.MM.YYYY'
        elif isinstance(special, int):
            ws.merge_cells(f'B{row}:E{row+special-1}')
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            row += special - 1
        
        row += 1
    
    # Approved By (Level 3 - CISO)
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'LEVEL 3: APPROVED BY (CISO / SECURITY DIRECTOR)'
    apply_style(ws[f'A{row}'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    approval_fields = [
        ('Name:', None),
        ('Approval Date:', 'date'),
        ('Approval Decision:', 'approval_dropdown'),
        ('Conditions/Notes:', 3),  # Multi-line
    ]
    
    for field, special in approval_fields:
        ws[f'A{row}'] = field
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        apply_style(cell, fill=INPUT_FILL)
        
        if special == 'approval_dropdown':
            validations['approval_decision'].add(cell)
        elif special == 'date':
            cell.number_format = 'DD.MM.YYYY'
        elif isinstance(special, int):
            ws.merge_cells(f'B{row}:E{row+special-1}')
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            row += special - 1
        
        row += 1
    
    # Next Review Details
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'NEXT REVIEW DETAILS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    row += 1
    next_review_fields = [
        ('Next Annual Review:', 'auto_365'),
        ('Review Responsible:', None),
        ('Special Considerations:', None),
    ]
    
    for field, special in next_review_fields:
        ws[f'A{row}'] = field
        apply_style(ws[f'A{row}'], font=BOLD_FONT)
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        apply_style(cell, fill=INPUT_FILL)
        
        if special == 'auto_365':
            cell.value = '=TODAY()+365'
            cell.number_format = 'DD.MM.YYYY'
            cell.font = Font(italic=True, color='808080')
        
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20
    
    return ws

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Generate complete BC/DR Consolidated Dashboard workbook"""
    
    try:
        logger.info(f"\n{'='*70}")
        logger.info(f"GENERATING: {WORKBOOK_TITLE}")
        logger.info(f"{'='*70}")
        logger.info(f"Version: {VERSION}")
        logger.info(f"Controls: {CONTROLS}")
        logger.info(f"Assessment ID: {ASSESSMENT_ID}")
        logger.info(f"{'='*70}\n")
        
        # Check for required normalized workbooks
        logger.info("Checking for required normalized workbooks...\n")
        missing_workbooks = []
        for key, filename in REQUIRED_WORKBOOKS.items():
            if not Path(filename).exists():
                missing_workbooks.append(filename)
                logger.info(f"  ❌ Missing: {filename}")
            else:
                file_size = Path(filename).stat().st_size
                logger.info(f"  ✅ Found: {filename} ({file_size:,} bytes)")
        
        if missing_workbooks:
            logger.info(f"\n{'='*70}")
            logger.error("{XMARK} ERROR: Missing Required Workbooks")
            logger.info(f"{'='*70}")
            logger.info("\nThe following normalized assessment workbooks are required:")
            for filename in missing_workbooks:
                logger.info(f"  • {filename}")
            logger.info("\nPREREQUISITE:")
            logger.info("  1. Generate all 4 assessment workbooks (Scripts 1-4)")
            logger.info("  2. Run: python normalize_bcdr_assessments.py")
            logger.info("  3. Then run this dashboard script")
            logger.info(f"\n{'='*70}\n")
            sys.exit(1)
        
        logger.info(f"\n✅ All 4 required workbooks found\n")
        
        # Load data from normalized workbooks
        logger.info("Loading data from assessment workbooks...")
        data = load_assessment_data()
        
        if data['missing']:
            logger.info("\n⚠️  WARNING: Some workbooks could not be loaded:")
            for wb_file in data['missing']:
                logger.info(f"  • {wb_file}")
            logger.info("\nProceeding with available data...\n")
        else:
            logger.info("{CHECK} All workbooks loaded successfully\n")
        
        # Calculate maturity score
        logger.info("Calculating BC/DR maturity score...")
        maturity_results = calculate_maturity_score(data)
        logger.info("{CHECK} Maturity Score: {maturity_results[0]:.1f}% ({maturity_results[2]})\n")
        
        # Create workbook
        logger.info("Creating dashboard worksheets...")
        wb = Workbook()
        wb.remove(wb.active)
        
        create_instructions_sheet(wb)
        logger.info("  ✅ Instructions")
        
        create_executive_dashboard_sheet(wb, data, maturity_results)
        logger.info("  ✅ Executive_Dashboard")
        
        create_detailed_metrics_sheet(wb, data)
        logger.info("  ✅ Detailed_Metrics")
        
        create_gap_summary_sheet(wb, data)
        logger.info("  ✅ Gap_Summary")
        
        create_evidence_checklist_sheet(wb, data)
        logger.info("  ✅ Evidence_Checklist")
        
        create_approval_signoff(wb)
        logger.info("  ✅ Approval_Sign_Off")
        
        # Save workbook
        filename = f"ISMS-IMP-A.5.30.S5_Compliance_Dashboard_{GENERATED_TIMESTAMP}.xlsx"
        wb.save(filename)
        
        # Summary
        logger.info(f"\n{'='*70}")
        logger.info("DASHBOARD GENERATED SUCCESSFULLY")
        logger.info(f"{'='*70}")
        logger.info(f"Filename: {filename}")
        logger.info(f"Worksheets: {len(wb.sheetnames)}")
        logger.info("\nWorksheet Details:")
        logger.info("  • Instructions (comprehensive usage guide)")
        logger.info("  • Executive_Dashboard (maturity score and key metrics)")
        logger.info("  • Detailed_Metrics (all metrics from 4 workbooks)")
        logger.info("  • Gap_Summary (consolidated critical gaps)")
        logger.info("  • Evidence_Checklist (evidence completeness)")
        logger.info("  • Approval_Sign_Off (3-level workflow)")
        logger.info(f"\n{'='*70}")
        logger.info("KEY METRICS:")
        logger.info(f"  • Overall BC/DR Maturity Score: {maturity_results[0]:.1f}%")
        logger.info(f"  • Maturity Level: {maturity_results[2]}")
        logger.info(f"  • Backup Coverage Score: {maturity_results[4]['backup']:.1f}/25")
        logger.info(f"  • Redundancy Score: {maturity_results[4]['redundancy']:.1f}/25")
        logger.info(f"  • RPO/RTO Alignment Score: {maturity_results[4]['rporto']:.1f}/25")
        logger.info(f"  • Testing Compliance Score: {maturity_results[4]['testing']:.1f}/25")
        logger.info(f"{'='*70}\n")
        
    except Exception as e:
        logger.info(f"\n{'='*70}")
        logger.error("{XMARK} ERROR: Failed to generate dashboard")
        logger.info(f"{'='*70}")
        logger.error(f"Error details: {str(e)}")
        logger.error(f"Error type: {type(e).__name__}")
        import traceback
        logger.info(f"\nFull traceback:")
        traceback.print_exc()
        logger.info(f"{'='*70}\n")
        sys.exit(1)

if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-01-31
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
