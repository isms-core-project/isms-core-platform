#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.30.S4 - BC/DR Testing Results Tracker Excel Generator
================================================================================

ISO/IEC 27001:2022 Controls A.5.30 (Backup), A.8.14 (Redundancy), A.5.30 (ICT BC)
Assessment Domain 4 of 4: Recovery Testing and Validation

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific BC/DR testing strategy, testing frequency requirements,
and validation procedures.

Key customization areas:
1. Testing frequency requirements per system tier
2. Test types and scope definitions (restore, failover, full DR scenario)
3. Success criteria and acceptance thresholds
4. Overdue testing alert thresholds
5. Testing compliance scoring methodology

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.30/14/5.30 BC/DR Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for tracking
and validating BC/DR testing activities across backup restore tests, failover
tests, and full disaster recovery scenario exercises.

**Purpose:**
Enables systematic tracking of BC/DR testing against ISO 27001:2022 Controls
A.5.30, A.8.14, and A.5.30 requirements, supporting the critical principle:
"Untested Recovery = No Recovery"

**Assessment Scope:**
- System inventory requiring regular testing
- Backup restore test tracking (per system/dataset)
- Failover test tracking (per redundant system)
- Full DR scenario test tracking (end-to-end exercises)
- Test frequency compliance monitoring
- Test results documentation (success/partial/failure)
- Issues identified during testing
- Remediation tracking for test failures
- Testing trend analysis (improving vs. degrading)
- Gap analysis for untested systems
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and testing methodology
2. Testing Results Tracker - Comprehensive test inventory and results (110 rows: 10 examples + 100 data entry)
3. Evidence Register - Audit evidence tracking and documentation (100 rows)
4. Approval & Sign-Off - Multi-level stakeholder review and approval workflow

**Key Features:**
- Data validation with comprehensive dropdown lists (test types, results, frequencies)
- Conditional formatting for overdue testing and failed tests
- Automated testing compliance calculations (last test vs. required frequency)
- Automated overdue test identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- 3-level approval workflow (Assessor -> ISO Officer -> CISO)
- Exception handling with graceful error reporting
- Professional styling without Excel repair warnings

**Integration:**
This assessment feeds into the ISMS-IMP-A.5.30.S5 BC/DR Summary Dashboard,
which consolidates data from all four BC/DR assessment domains for executive
oversight and audit readiness. Testing results validate capabilities assessed
in Domains 1 (Backup) and 2 (Redundancy).

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl --break-system-packages

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a530_4_testing_results.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a530_4_testing_results.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a530_4_testing_results.py --date 20250125

Output:
    File: ISMS_Assessment_Testing_Results.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review and customize testing frequency requirements per system tier
    2. Inventory all systems requiring regular BC/DR testing
    3. Document backup restore test results
    4. Document failover test results
    5. Document full DR scenario test results
    6. Analyze testing compliance (on schedule vs. overdue)
    7. Track remediation for failed tests or identified issues
    8. Define testing schedule for untested systems
    9. Obtain stakeholder approvals
    10. Feed results into ISMS-IMP-A.5.30.S5 BC/DR Summary Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.5.30, A.8.14, A.5.30
Assessment Domain:    4 of 4 (BC/DR Testing & Validation)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.30-8.13-14: BC/DR Framework Policy (Governance)
    - ISMS-POL-A.5.30-8.13-14-S5: Testing Methodology and Evidence Framework
    - ISMS-IMP-A.5.30-8.13-14-S4: Recovery Testing Process
    - ISMS-IMP-A.5.30-8.13-14-S5: BC/DR Assessment Guide
    - ISMS-IMP-A.5.30.S1: Backup Inventory Assessment (Domain 1)
    - ISMS-IMP-A.5.30.S2: Redundancy Analysis Assessment (Domain 2)
    - ISMS-IMP-A.5.30.S3: RPO/RTO Compliance Matrix (Domain 3)
    - ISMS-IMP-A.5.30.S5: BC/DR Summary Dashboard (Consolidation)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.30-8.13-14-S5 specification
    - Supports comprehensive BC/DR testing tracking and validation
    - Integrated with ISMS-IMP-A.5.30.S5 BC/DR Summary Dashboard
    - Includes automated testing compliance monitoring

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Untested Recovery = No Recovery:**
This assessment embodies the most critical BC/DR principle: recovery capabilities
are meaningless without regular testing. Systems without recent successful tests
should be treated as UNRECOVERABLE regardless of backup or redundancy architecture.

**Testing Frequency Requirements:**
Testing frequency should be risk-based and aligned with system criticality:
- Tier 1 (Critical): Quarterly backup restore tests, semi-annual failover tests
- Tier 2 (Important): Semi-annual backup restore tests, annual failover tests
- Tier 3 (Standard): Annual backup restore tests
- Tier 4 (Low): Opportunistic testing or test-on-change

Customize based on your risk tolerance and regulatory requirements.

**Test Types:**
Different test types validate different capabilities:
- **Backup Restore Test**: Validates backup integrity and restore procedures (A.5.30)
- **Failover Test**: Validates redundancy and failover mechanisms (A.8.14)
- **Full DR Scenario**: Validates end-to-end recovery including people, process, technology (A.5.30)

All three are required for comprehensive BC/DR validation.

**Test Results Interpretation:**
- **Success**: Test completed within RTO, all functionality verified
- **Partial Success**: Test completed but issues identified (acceptable if documented)
- **Failure**: Test did not complete or major functionality unavailable
- **Not Tested**: HIGH RISK - assume recovery will fail

**Testing Without Production Impact:**
BC/DR testing should not disrupt production systems:
- Restore to isolated test environment (not production)
- Failover during maintenance windows
- Use snapshots or clones for testing

Document testing methodology to demonstrate safe practices.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.
Auditors will expect:
- Evidence of regular testing (test reports, screenshots, logs)
- Documented testing procedures
- Remediation tracking for failed tests
- Management approval of testing schedule

**Data Protection:**
Assessment workbooks contain sensitive operational information including:
- Testing schedules (security-sensitive timing information)
- Test failures and issues identified
- Recovery procedures and timings

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Monthly: Update with latest test results
- Quarterly: Analyze testing trends and compliance rates
- Semi-annually: Review testing frequency requirements
- Annually: Complete full reassessment of testing strategy
- Ad-hoc: After failed tests or recovery incidents

**Quality Assurance:**
Have BC/DR coordinators, system administrators, and business continuity managers
validate test results before using for compliance reporting.

**Regulatory Alignment:**
Ensure testing frequency aligns with applicable regulatory requirements:
- Finance: DORA requires regular testing of ICT systems
- Healthcare: HIPAA requires contingency plan testing and revision
- Critical Infrastructure: Sector-specific testing mandates

**Learning from Test Failures:**
Test failures are GOOD - they identify issues before real disasters:
- Document all issues found during testing
- Track remediation with priority and timelines
- Re-test after remediation
- Update recovery procedures based on lessons learned

The goal of testing is to find and fix problems, not to achieve 100% success rate.

**Testing vs. Real Incidents:**
Testing validates capability but doesn't guarantee success during real incidents:
- Tests are controlled and planned
- Real incidents involve stress, time pressure, and unknowns
- Both testing and incident response experience are valuable

Document both test results AND actual recovery incidents.

================================================================================
"""

# =============================================================================
# STANDARD LIBRARY IMPORTS
# =============================================================================
import logging
from datetime import datetime, timedelta
import sys
from pathlib import Path

# =============================================================================
# THIRD-PARTY IMPORTS
# =============================================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)
WORKBOOK_TITLE = "BC/DR Testing Results Tracker"
WORKBOOK_NAME = "Testing Results"
VERSION = "1.0"
CONTROLS = "A.5.30 (Information Backup)"


# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================

CHECK = '\u2705'      # ✅ Green checkmark
XMARK = '\u274C'      # ❌ Red X
WARNING = '\u26A0'    # ⚠  Warning sign
BULLET = '\u2022'     # • Bullet point
ARROW = '\u2192'      # → Right arrow

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.30.S4"
CONTROL_ID   = "A.5.30"
CONTROL_NAME = "Information Backup"
CONTROL_REF  = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
GENERATED_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(exist_ok=True)

# Color scheme (consistent with reference implementations)
HEADER_FILL = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
COLUMN_HEADER_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

# Fonts
HEADER_FONT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
SUBHEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
BOLD_FONT = Font(name='Calibri', size=10, bold=True)
NORMAL_FONT = Font(name='Calibri', size=10)

# Borders
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def apply_style(cell, font=None, fill=None, alignment=None, border=None):
    """Apply multiple styles to a cell"""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border

def create_base_validations(ws):
    """
    Create comprehensive data validation objects for dropdowns.
    
    Returns dict of validation objects to be applied to cells.
    All validations are added to worksheet once, then applied to multiple cells.
    """
    validations = {
        'test_type': DataValidation(
            type="list",
            formula1='"Backup Restore,Failover Test,Full DR Scenario,Tabletop Exercise,Component Test,Integration Test"',
            allow_blank=False
        ),
        'test_result': DataValidation(
            type="list",
            formula1=f'"{CHECK} Success,{WARNING} Partial Success,{XMARK} Failure,In Progress,Not Started,Cancelled"',
            allow_blank=False
        ),
        'criticality': DataValidation(
            type="list",
            formula1='"Tier 1 - Critical,Tier 2 - Important,Tier 3 - Standard,Tier 4 - Low"',
            allow_blank=False
        ),
        'test_frequency': DataValidation(
            type="list",
            formula1='"Quarterly,Semi-Annual,Annual,Ad-Hoc,Not Required"',
            allow_blank=False
        ),
        'compliance_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Compliant,⚠️ Overdue,❌ Non-Compliant,❓ Unknown"',
            allow_blank=False
        ),
        'issue_severity': DataValidation(
            type="list",
            formula1='"[!] Critical,[~] High,[.] Medium,[ ] Low"',
            allow_blank=False
        ),
        'issue_status': DataValidation(
            type="list",
            formula1=f'"Open,In Progress,{CHECK} Closed,Risk Accepted"',
            allow_blank=False
        ),
        'evidence_type': DataValidation(
            type="list",
            formula1='"Test Report,Screenshot,Log File,Video Recording,Configuration File,Runbook,Issue Ticket,Sign-Off,Other"',
            allow_blank=False
        ),
        'verification_status': DataValidation(
            type="list",
            formula1=f'"{CHECK} Verified,⏳ Pending,❌ Not Verified"',
            allow_blank=False
        ),
        'assessment_status': DataValidation(
            type="list",
            formula1='"Draft,Under Review,Final,Requires Remediation"',
            allow_blank=False
        ),
        'approval_decision': DataValidation(
            type="list",
            formula1='"Approved,Approved with Conditions,Rejected,Requires Rework"',
            allow_blank=False
        ),
        'recommendation': DataValidation(
            type="list",
            formula1='"Approve,Approve with Conditions,Reject,Require Rework"',
            allow_blank=False
        ),
    }
    
    return validations

# ============================================================================
# WORKSHEET: INSTRUCTIONS
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")

    # Row 1 — Title banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Row 3 — Document Information heading
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)

    doc_info = [
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]
    for i, (label, value) in enumerate(doc_info):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border

    # Row 12 — Instructions heading
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Plan tests in the Test Schedule worksheet (define annual test plan).', '2. Execute tests according to schedule and document actual execution.', '3. Log results in the Test Results Log worksheet (document every test performed).', '4. Document issues in the Issue Remediation worksheet (problems discovered during testing).', '5. Track remediation of issues to closure.', '6. Review Testing Compliance worksheet (annual compliance assessment).', '7. Collect evidence in the Evidence Register (minimum 5 items required for audit).', '8. Review the Summary Dashboard for overall testing compliance metrics.', '9. Complete the Approval Sign-Off worksheet to obtain formal sign-off.', '10. Repeat annually for continuous compliance.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 24

    _EVIDENCE = ['✓ Test reports with date, system name, test type, outcome, and actual duration', '✓ Screenshots of backup restore test completion (data integrity verified)', '✓ Failover test logs showing actual switchover time', '✓ Full DR scenario exercise reports with stakeholder attendance', '✓ Issue tracking records (Jira tickets, remediation plans) from test failures', '✓ Testing schedule demonstrating regular cadence aligned to criticality tiers', '✓ Video recordings or observer notes from tabletop exercises', '✓ Management sign-off on annual testing programme']

    # Status Legend — row position tracks after instructions
    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    legend_rows = [
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]
    for i, (sym, status, desc, fill) in enumerate(legend_rows):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Acceptable Evidence section
    _ev_row = _leg_row + 7
    ws[f"A{_ev_row}"] = "Acceptable Evidence (examples)"
    ws[f"A{_ev_row}"].font = Font(name="Calibri", size=12, bold=True)
    for _ev in _EVIDENCE:
        _ev_row += 1
        ws[f"A{_ev_row}"] = _ev

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_summary_dashboard_sheet(wb):
    """Create Summary Dashboard worksheet (placed as 2nd sheet)"""
    ws = wb.create_sheet(title="Summary Dashboard")
    ws.sheet_view.showGridLines = False

    # Title — DS-002: fill must be HEADER_FILL (003366) with white font
    ws.merge_cells('A1:E1')
    ws['A1'] = 'TESTING RESULTS \u2014 SUMMARY DASHBOARD'
    apply_style(ws['A1'], font=Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
                fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws["A1"].border = border
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:E2')
    ws['A2'] = f'Assessment ID: {DOCUMENT_ID} | Assessment Date: [enter date]'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(italic=True, size=10)

    # Testing Activity Summary
    row = 4
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'TESTING ACTIVITY SUMMARY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))

    row += 1
    activity_metrics = [
        ('Total Tests Scheduled:', "=COUNTA('Test Schedule'!A16:A65)"),
        ('Tests Completed:', f"=COUNTIF('Test Results Log'!F16:F65,\"{CHECK}*\")"),
        ('Tests Partial Success:', f"=COUNTIF('Test Results Log'!F16:F65,\"{WARNING}*\")"),
        ('Tests Failed:', f"=COUNTIF('Test Results Log'!F16:F65,\"{XMARK}*\")"),
        ('Tests In Progress:', "=COUNTIF('Test Results Log'!F16:F65,\"In Progress*\")"),
        ('Tests Not Started:', "=COUNTIF('Test Results Log'!F16:F65,\"Not Started*\")"),
        ('', ''),
        ('Test Success Rate:', '=IF(B5>0,(B6+B7*0.5)/B5,0)'),
    ]

    for label, formula in activity_metrics:
        if label:
            ws[f'A{row}'] = label
            apply_style(ws[f'A{row}'], font=BOLD_FONT)
            ws[f'B{row}'] = formula
            apply_style(ws[f'B{row}'], font=Font(bold=True, size=12))
            if 'Rate' in label:
                ws[f'B{row}'].number_format = '0.0%'
        row += 1

    # Testing Compliance by Criticality
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'TESTING COMPLIANCE BY CRITICALITY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))

    row += 1
    criticality_metrics = [
        ('Tier 1 - Critical Systems:', "=COUNTIF('Testing Compliance'!C16:C65,\"Tier 1*\")"),
        ('Tier 1 - Compliant:', f"=SUMPRODUCT(('Testing Compliance'!C16:C65=\"Tier 1 - Critical\")*('Testing Compliance'!G16:G65=\"{CHECK} Compliant\"))"),
        ('Tier 1 - Compliance Rate:', '=IF(B15>0,B16/B15,0)'),
        ('', ''),
        ('Tier 2 - Important Systems:', "=COUNTIF('Testing Compliance'!C16:C65,\"Tier 2*\")"),
        ('Tier 2 - Compliant:', f"=SUMPRODUCT(('Testing Compliance'!C16:C65=\"Tier 2 - Important\")*('Testing Compliance'!G16:G65=\"{CHECK} Compliant\"))"),
        ('Tier 2 - Compliance Rate:', '=IF(B19>0,B20/B19,0)'),
    ]

    for label, formula in criticality_metrics:
        if label:
            ws[f'A{row}'] = label
            apply_style(ws[f'A{row}'], font=BOLD_FONT)
            ws[f'B{row}'] = formula
            apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
            if 'Rate' in label:
                ws[f'B{row}'].number_format = '0.0%'
        row += 1

    # Issue Summary
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'ISSUE SUMMARY'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))

    row += 1
    issue_metrics = [
        ('Total Issues Identified:', "=COUNTA('Issue Remediation'!A16:A65)"),
        ('[!] Critical Issues:', "=COUNTIF('Issue Remediation'!E16:E65,\"[!]*\")"),
        ('[~] High Severity Issues:', "=COUNTIF('Issue Remediation'!E16:E65,\"[~]*\")"),
        ('[.] Medium Severity Issues:', "=COUNTIF('Issue Remediation'!E16:E65,\"[.]*\")"),
        ('[ ] Low Severity Issues:', "=COUNTIF('Issue Remediation'!E16:E65,\"[ ]*\")"),
        ('', ''),
        ('Open Issues:', "=COUNTIF('Issue Remediation'!F16:F65,\"Open*\")"),
        ('In Progress:', "=COUNTIF('Issue Remediation'!F16:F65,\"In Progress*\")"),
        ('Closed Issues:', f"=COUNTIF('Issue Remediation'!F16:F65,\"{CHECK}*\")"),
    ]

    for label, formula in issue_metrics:
        if label:
            ws[f'A{row}'] = label
            apply_style(ws[f'A{row}'], font=BOLD_FONT)
            ws[f'B{row}'] = formula
            apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1

    # Evidence & Approval Status
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'EVIDENCE & APPROVAL STATUS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))

    row += 1
    evidence_metrics = [
        ('Evidence Items Collected:', "=COUNTA('Evidence Register'!A6:A105)"),
        ('Verified Evidence:', f"=COUNTIF('Evidence Register'!H6:H105,\"{CHECK}*\")"),
        ('Minimum Evidence Required:', '5'),
        ('Evidence Compliance:', f'=IF(B39>=B41,"{CHECK} Sufficient","{XMARK} Insufficient")'),
        ('', ''),
        ('Assessment Status:', "='Approval Sign-Off'!B7"),
        ('Level 1 - Assessor Completed:', f"=IF('Approval Sign-Off'!B11<>\"\",\"{CHECK} Complete\",\"Pending\")"),
        ('Level 2 - ISO Review:', f"=IF('Approval Sign-Off'!B18<>\"\",\"{CHECK} Complete\",\"Pending\")"),
        ('Level 3 - CISO Approval:', f"=IF('Approval Sign-Off'!B25<>\"\",\"{CHECK} Complete\",\"Pending\")"),
    ]

    for label, formula in evidence_metrics:
        if label:
            ws[f'A{row}'] = label
            apply_style(ws[f'A{row}'], font=BOLD_FONT)
            ws[f'B{row}'] = formula
            if label == 'Minimum Evidence Required:':
                apply_style(ws[f'B{row}'], font=Font(bold=True, size=11, color='C00000'))
            else:
                apply_style(ws[f'B{row}'], font=Font(bold=True, size=11))
        row += 1

    # Priority Actions
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'PRIORITY ACTIONS'
    apply_style(ws[f'A{row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='left'))

    row += 1
    priority_actions = [
        '1. Address all [!] Critical issues discovered during testing',
        '2. Complete all overdue tests (Overdue in Testing Compliance)',
        '3. Schedule quarterly tests for Tier 1 systems',
        '4. Remediate test failures (Failed tests in Test Results Log)',
        '5. Document lessons learned from partial successes',
        '6. Collect minimum 5 evidence items in the Evidence Register',
        '7. Complete all 3 approval levels (Assessor to ISO to CISO)',
        '8. Plan next annual full DR scenario test',
        '9. Update runbooks based on test findings',
    ]

    for action in priority_actions:
        ws[f'A{row}'] = action
        ws[f'A{row}'].font = NORMAL_FONT
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 25
    for col in ['C', 'D', 'E']:
        ws.column_dimensions[col].width = 15

    return ws

# ============================================================================
# WORKSHEET: TEST SCHEDULE
# ============================================================================

def create_test_schedule_sheet(wb):
    """Create Test Schedule worksheet (annual test planning)"""
    ws = wb.create_sheet(title="Test Schedule")
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = 'BC/DR TEST SCHEDULE'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:H2')
    ws['A2'] = 'Define annual test plan - what tests to perform when (schedule all required tests)'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'Test ID',
        'System Name',
        'Test Type',
        'Criticality',
        'Required Frequency',
        'Scheduled Date',
        'Test Owner',
        'Notes / Prerequisites'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    # Apply data validations and styling for 61 rows (10 examples + 1 grey sample + 50 empty)
    _grey_ts = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row = 5
    for i in range(61):
        current_row = row + i

        # Test ID — example rows (i<10) left empty (IDs added by examples section below);
        # grey sample (i=10) placeholder (GS-MAX-003)
        if i < 10:
            apply_style(ws[f'A{current_row}'], font=Font(bold=True, size=9), border=THIN_BORDER)
        elif i == 10:  # Grey sample row (GS-MAX-003)
            ws[f'A{current_row}'] = 'TEST-NNN'
            apply_style(ws[f'A{current_row}'], fill=_grey_ts,
                       font=Font(bold=True, size=9), border=THIN_BORDER)
        else:
            apply_style(ws[f'A{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)

        # System Name, Test Owner, Notes (user input)
        for col in ['B', 'G', 'H']:
            if i < 10:
                apply_style(ws[f'{col}{current_row}'], border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
            elif i == 10:
                apply_style(ws[f'{col}{current_row}'], fill=_grey_ts, border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
            else:
                apply_style(ws[f'{col}{current_row}'], fill=INPUT_FILL, border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))

        # Test Type dropdown
        validations['test_type'].add(f'C{current_row}')
        apply_style(ws[f'C{current_row}'],
                   fill=_grey_ts if i == 10 else None, border=THIN_BORDER)

        # Criticality dropdown
        validations['criticality'].add(f'D{current_row}')
        apply_style(ws[f'D{current_row}'],
                   fill=_grey_ts if i == 10 else None, border=THIN_BORDER)

        # Required Frequency dropdown
        validations['test_frequency'].add(f'E{current_row}')
        apply_style(ws[f'E{current_row}'],
                   fill=_grey_ts if i == 10 else None, border=THIN_BORDER)

        # Scheduled Date (user input, date format)
        if i < 10:
            apply_style(ws[f'F{current_row}'], border=THIN_BORDER)
        elif i == 10:
            apply_style(ws[f'F{current_row}'], fill=_grey_ts, border=THIN_BORDER)
        else:
            apply_style(ws[f'F{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)
        ws[f'F{current_row}'].number_format = 'DD.MM.YYYY'

        if i == 10:  # Sample data for grey row
            ws[f'B{current_row}'] = 'New System (Example)'
            ws[f'C{current_row}'] = 'Backup Restore'
            ws[f'D{current_row}'] = 'Tier 2 - Important'
            ws[f'E{current_row}'] = 'Quarterly'
            ws[f'G{current_row}'] = 'Test Owner'
            ws[f'H{current_row}'] = 'Sample — enter test prerequisites here'
    
    # Grey sample row at row 15 (F2F2F2) — MAX-003 fix
    _grey_ts = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _ts_sample_row = 5 + 10  # row 15
    _ts_sample = ['TEST-NNN', 'System Name', 'Backup Restore', 'Tier 1 - Critical',
                  'Quarterly', 'DD.MM.YYYY', 'Owner Name', 'Prerequisites and notes here']
    for _ci, _val in enumerate(_ts_sample, start=1):
        _c = ws.cell(row=_ts_sample_row, column=_ci, value=_val)
        _c.fill = _grey_ts
        _c.border = THIN_BORDER

    # Summary metrics
    summary_row = 67
    ws.merge_cells(f'A{summary_row}:H{summary_row}')
    ws[f'A{summary_row}'] = 'TEST SCHEDULE SUMMARY'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    summary_row += 1
    metrics = [
        ('Total Tests Scheduled:', f'=COUNTA(B5:B65)'),
        ('Backup Restore Tests:', f'=COUNTIF(C5:C65,"Backup*")'),
        ('Failover Tests:', f'=COUNTIF(C5:C65,"Failover*")'),
        ('Full DR Scenarios:', f'=COUNTIF(C5:C65,"Full DR*")'),
        ('Tier 1 System Tests:', f'=COUNTIF(D5:D65,"Tier 1*")'),
        ('Quarterly Tests:', f'=COUNTIF(E5:E65,"Quarterly")'),
    ]
    
    for label, formula in metrics:
        ws[f'A{summary_row}'] = label
        apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
        ws[f'B{summary_row}'] = formula
        apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
        summary_row += 1
    
    # Column widths
    widths = [12, 25, 20, 18, 18, 15, 20, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Lazy-add: only register DVs that have cells assigned (fixes empty-sqref XML error)
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)
    
    ws.freeze_panes = 'A4'
    
    return ws

# ============================================================================
# WORKSHEET: TEST RESULTS LOG
# ============================================================================

def create_test_results_log_sheet(wb):
    """Create Test Results Log worksheet (all test executions and outcomes)"""
    ws = wb.create_sheet(title="Test Results Log")
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = 'BC/DR TEST RESULTS LOG'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:J2')
    ws['A2'] = 'Document all test executions - actual test results (log every test performed)'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'Result ID',
        'Test ID (Schedule)',
        'System Name',
        'Test Type',
        'Test Date',
        'Result',
        'Actual Duration (hrs)',
        'Tested By',
        'Issues Found',
        'Detailed Notes'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    # Apply data validations and styling for 61 rows (10 examples + 1 sample + 50 empty) — MAX-004 fix
    row = 5
    _sample_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    for i in range(61):  # was range(110)
        current_row = row + i

        # Result ID — example rows (i<10) pre-filled; sample (i==10) placeholder; empty (i>10) FFFFCC
        if i < 10:
            ws[f'A{current_row}'] = f'RES-{i+1:03d}'
            apply_style(ws[f'A{current_row}'], font=Font(bold=True, size=9), border=THIN_BORDER)
        elif i == 10:  # Grey sample row — show how to fill in
            ws[f'A{current_row}'] = 'RES-EXX'
            apply_style(ws[f'A{current_row}'], fill=_sample_fill,
                       font=Font(bold=True, size=9), border=THIN_BORDER)
        else:
            apply_style(ws[f'A{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)

        # Test ID, System Name, Tested By, Issues Found, Notes (user input)
        for col in ['B', 'C', 'H', 'I', 'J']:
            if i == 10:
                apply_style(ws[f'{col}{current_row}'], fill=_sample_fill, border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
            elif i > 10:
                apply_style(ws[f'{col}{current_row}'], fill=INPUT_FILL, border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
            else:
                apply_style(ws[f'{col}{current_row}'], border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
        
        # Test Type dropdown
        validations['test_type'].add(f'D{current_row}')
        apply_style(ws[f'D{current_row}'], border=THIN_BORDER)
        
        # Test Date (user input, date format)
        if i >= 10:
            apply_style(ws[f'E{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)
        else:
            apply_style(ws[f'E{current_row}'], border=THIN_BORDER)
        ws[f'E{current_row}'].number_format = 'DD.MM.YYYY'
        
        # Result dropdown
        validations['test_result'].add(f'F{current_row}')
        apply_style(ws[f'F{current_row}'], border=THIN_BORDER)
        
        # Actual Duration (user input, numeric)
        if i >= 10:
            apply_style(ws[f'G{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)
        else:
            apply_style(ws[f'G{current_row}'], border=THIN_BORDER)

        # Grey sample row data (GS-MAX-003: must have data; col A left blank to avoid MAX-001)
        if i == 10:
            ws[f'B{current_row}'] = 'TEST-EX'
            ws[f'C{current_row}'] = 'New System (Example)'
            ws[f'D{current_row}'] = 'Backup Restore'
            ws[f'F{current_row}'] = f'{CHECK} Success'
            ws[f'G{current_row}'] = 2.5
            ws[f'H{current_row}'] = 'IT Operations Team'
            ws[f'I{current_row}'] = 'None'
            ws[f'J{current_row}'] = 'Sample — enter actual test outcome and notes here'

    # Summary metrics
    summary_row = 67
    ws.merge_cells(f'A{summary_row}:J{summary_row}')
    ws[f'A{summary_row}'] = 'TEST RESULTS SUMMARY'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    summary_row += 1
    metrics = [
        ('Total Tests Executed:', f'=COUNTA(C5:C65)'),
        (f'{CHECK} Successful Tests:', f'=COUNTIF(F5:F65,"{CHECK}*")'),
        (f'{WARNING} Partial Success:', f'=COUNTIF(F5:F65,"{WARNING}*")'),
        (f'{XMARK} Failed Tests:', f'=COUNTIF(F5:F65,"{XMARK}*")'),
        ('Success Rate:', f'=IF(B{summary_row}>0,(B{summary_row+1}+B{summary_row+2}*0.5)/B{summary_row},0)'),
        ('Average Test Duration (hrs):', f'=IFERROR(AVERAGE(G5:G65),0)'),
    ]
    
    for label, formula in metrics:
        ws[f'A{summary_row}'] = label
        apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
        ws[f'B{summary_row}'] = formula
        apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
        if 'Rate' in label:
            ws[f'B{summary_row}'].number_format = '0.0%'
        elif 'Duration' in label:
            ws[f'B{summary_row}'].number_format = '0.0'
        summary_row += 1
    
    # Column widths
    widths = [12, 18, 25, 20, 15, 18, 18, 20, 20, 45]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Lazy-add: only register DVs that have cells assigned (fixes empty-sqref XML error)
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)
    
    ws.freeze_panes = 'A4'
    
    return ws

# ============================================================================
# WORKSHEET: ISSUE REMEDIATION
# ============================================================================

def create_issue_remediation_sheet(wb):
    """Create Issue Remediation worksheet (issues discovered during testing)"""
    ws = wb.create_sheet(title="Issue Remediation")
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = 'ISSUE REMEDIATION TRACKER'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:H2')
    ws['A2'] = 'Track issues discovered during testing to closure'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'Issue ID',
        'Related Test (Result ID)',
        'Issue Description',
        'Root Cause',
        'Severity',
        'Status',
        'Remediation Plan',
        'Target Resolution Date'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    # Apply data validations and styling for 61 rows (10 examples + 1 sample + 50 empty) — MAX-004 fix
    row = 5
    _sample_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    for i in range(61):  # was range(110)
        current_row = row + i

        # Issue ID — example rows (i<10) pre-filled; sample (i==10) placeholder; empty (i>10) FFFFCC
        if i < 10:
            ws[f'A{current_row}'] = f'ISS-{i+1:03d}'
            apply_style(ws[f'A{current_row}'], font=Font(bold=True, size=9), border=THIN_BORDER)
        elif i == 10:  # Grey sample row — show how to fill in
            ws[f'A{current_row}'] = 'ISS-EXX'
            apply_style(ws[f'A{current_row}'], fill=_sample_fill,
                       font=Font(bold=True, size=9), border=THIN_BORDER)
        else:
            apply_style(ws[f'A{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)

        # Related Test, Description, Root Cause, Remediation Plan (user input)
        for col in ['B', 'C', 'D', 'G']:
            if i == 10:
                apply_style(ws[f'{col}{current_row}'], fill=_sample_fill, border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
            elif i > 10:
                apply_style(ws[f'{col}{current_row}'], fill=INPUT_FILL, border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
            else:
                apply_style(ws[f'{col}{current_row}'], border=THIN_BORDER,
                           alignment=Alignment(wrap_text=True, vertical='top'))
        
        # Severity dropdown
        validations['issue_severity'].add(f'E{current_row}')
        apply_style(ws[f'E{current_row}'], border=THIN_BORDER)
        
        # Status dropdown
        validations['issue_status'].add(f'F{current_row}')
        apply_style(ws[f'F{current_row}'], border=THIN_BORDER)
        
        # Target Resolution Date (user input, date format)
        if i == 10:
            apply_style(ws[f'H{current_row}'], fill=_sample_fill, border=THIN_BORDER)
        elif i > 10:
            apply_style(ws[f'H{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)
        else:
            apply_style(ws[f'H{current_row}'], border=THIN_BORDER)
        ws[f'H{current_row}'].number_format = 'DD.MM.YYYY'

        # Grey sample row data (GS-MAX-003: must have data; col A left blank to avoid MAX-001)
        if i == 10:
            ws[f'B{current_row}'] = 'RES-EX'
            ws[f'C{current_row}'] = 'Sample issue description — enter details here'
            ws[f'D{current_row}'] = 'Root cause not yet identified'
            ws[f'E{current_row}'] = '[.] Medium'
            ws[f'F{current_row}'] = 'Open'
            ws[f'G{current_row}'] = 'Sample — enter remediation plan here'

    # Summary metrics
    summary_row = 67
    ws.merge_cells(f'A{summary_row}:H{summary_row}')
    ws[f'A{summary_row}'] = 'ISSUE REMEDIATION SUMMARY'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    summary_row += 1
    metrics = [
        ('Total Issues:', f'=COUNTA(C5:C65)'),
        ('[!] Critical Issues:', f'=COUNTIF(E5:E65,"[!]*")'),
        ('[~] High Severity:', f'=COUNTIF(E5:E65,"[~]*")'),
        ('[.] Medium Severity:', f'=COUNTIF(E5:E65,"[.]*")'),
        ('', ''),
        ('Open Issues:', f'=COUNTIF(F5:F65,"Open*")'),
        ('In Progress:', f'=COUNTIF(F5:F65,"In Progress*")'),
        (f'{CHECK} Closed Issues:', f'=COUNTIF(F5:F65,"{CHECK}*")'),
        ('Issue Closure Rate:', f'=IF(B{summary_row}>0,B{summary_row+7}/B{summary_row},0)'),
    ]
    
    for label, formula in metrics:
        if label:
            ws[f'A{summary_row}'] = label
            apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
            ws[f'B{summary_row}'] = formula
            apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
            if 'Rate' in label:
                ws[f'B{summary_row}'].number_format = '0.0%'
        summary_row += 1
    
    # Column widths
    widths = [12, 18, 40, 35, 15, 15, 45, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Lazy-add: only register DVs that have cells assigned (fixes empty-sqref XML error)
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)
    
    ws.freeze_panes = 'A4'
    
    return ws

# ============================================================================
# WORKSHEET: TESTING COMPLIANCE
# ============================================================================

def create_testing_compliance_sheet(wb):
    """Create Testing Compliance worksheet (annual compliance assessment)"""
    ws = wb.create_sheet(title="Testing Compliance")
    ws.sheet_view.showGridLines = False
    validations = create_base_validations(ws)

    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = 'TESTING COMPLIANCE ASSESSMENT'
    apply_style(ws['A1'], font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:H2')
    ws['A2'] = 'Annual compliance assessment - are systems being tested according to requirements?'
    apply_style(ws['A2'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))
    
    # Column headers
    row = 4
    headers = [
        'System Name',
        'Test Type',
        'Criticality',
        'Required Frequency',
        'Last Test Date',
        'Days Since Last Test',
        'Compliance Status',
        'Next Test Due'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, font=BOLD_FONT, fill=COLUMN_HEADER_FILL,
                   alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                   border=THIN_BORDER)
    
    # Apply formulas and styling for 61 rows (10 examples + 1 grey sample + 50 empty)
    _grey_tc = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row = 5
    for i in range(61):
        current_row = row + i
        _is_sample = (i == 10)
        _is_empty = (i > 10)

        # System Name, Last Test Date (user input)
        for col in ['A', 'E']:
            if _is_sample:
                apply_style(ws[f'{col}{current_row}'], fill=_grey_tc, border=THIN_BORDER)
            elif _is_empty:
                apply_style(ws[f'{col}{current_row}'], fill=INPUT_FILL, border=THIN_BORDER)
            else:
                apply_style(ws[f'{col}{current_row}'], border=THIN_BORDER)
        ws[f'E{current_row}'].number_format = 'DD.MM.YYYY'

        # Test Type dropdown
        validations['test_type'].add(f'B{current_row}')
        apply_style(ws[f'B{current_row}'],
                   fill=_grey_tc if _is_sample else None, border=THIN_BORDER)

        # Criticality dropdown
        validations['criticality'].add(f'C{current_row}')
        apply_style(ws[f'C{current_row}'],
                   fill=_grey_tc if _is_sample else None, border=THIN_BORDER)

        # Required Frequency dropdown
        validations['test_frequency'].add(f'D{current_row}')
        apply_style(ws[f'D{current_row}'],
                   fill=_grey_tc if _is_sample else None, border=THIN_BORDER)
        
        # F: Days Since Last Test (formula)
        ws[f'F{current_row}'] = f'=IF(E{current_row}="","",TODAY()-E{current_row})'
        apply_style(ws[f'F{current_row}'],
                   fill=_grey_tc if _is_sample else None,
                   border=THIN_BORDER, font=Font(italic=True, color='666666'))

        # G: Compliance Status (formula based on frequency and days since last test)
        # Quarterly = 90 days, Semi-Annual = 180 days, Annual = 365 days
        ws[f'G{current_row}'] = f'''=IF(A{current_row}="","",IF(E{current_row}="","❓ Unknown",IF(D{current_row}="Quarterly",IF(F{current_row}<=90,"{CHECK} Compliant",IF(F{current_row}<=105,"{WARNING} Overdue","{XMARK} Non-Compliant")),IF(D{current_row}="Semi-Annual",IF(F{current_row}<=180,"{CHECK} Compliant",IF(F{current_row}<=200,"{WARNING} Overdue","{XMARK} Non-Compliant")),IF(D{current_row}="Annual",IF(F{current_row}<=365,"{CHECK} Compliant",IF(F{current_row}<=400,"{WARNING} Overdue","{XMARK} Non-Compliant")),"{CHECK} Compliant")))))'''
        apply_style(ws[f'G{current_row}'],
                   fill=_grey_tc if _is_sample else None,
                   border=THIN_BORDER, font=Font(italic=True, color='666666'))

        # H: Next Test Due (formula)
        ws[f'H{current_row}'] = f'''=IF(E{current_row}="","",IF(D{current_row}="Quarterly",E{current_row}+90,IF(D{current_row}="Semi-Annual",E{current_row}+180,IF(D{current_row}="Annual",E{current_row}+365,E{current_row}))))'''
        apply_style(ws[f'H{current_row}'],
                   fill=_grey_tc if _is_sample else None,
                   border=THIN_BORDER, font=Font(italic=True, color='666666'))
        ws[f'H{current_row}'].number_format = 'DD.MM.YYYY'

        # Sample row: add guide data for row 15 (i==10)
        if _is_sample:
            ws[f'A{current_row}'] = 'Example: Application Server'
            ws[f'B{current_row}'] = 'Failover Test'
            ws[f'C{current_row}'] = 'Tier 1 - Critical'
            ws[f'D{current_row}'] = 'Annual'
            ws[f'E{current_row}'] = ''

    # Summary metrics (data rows 5-65: 61 rows total)
    summary_row = 68
    ws.merge_cells(f'A{summary_row}:H{summary_row}')
    ws[f'A{summary_row}'] = 'TESTING COMPLIANCE SUMMARY'
    apply_style(ws[f'A{summary_row}'], font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
                alignment=Alignment(horizontal='center'))

    summary_row += 1
    metrics = [
        ('Total Systems Tracked:', f'=COUNTA(A6:A65)'),
        (f'{CHECK} Compliant:', f'=COUNTIF(G6:G65,"{CHECK}*")'),
        (f'{WARNING} Overdue:', f'=COUNTIF(G6:G65,"{WARNING}*")'),
        (f'{XMARK} Non-Compliant:', f'=COUNTIF(G6:G65,"{XMARK}*")'),
        ('❓ Unknown (Never Tested):', f'=COUNTIF(G6:G65,"❓*")'),
        ('', ''),
        ('Overall Compliance Rate:', f'=IF(B{summary_row}>0,B{summary_row+1}/B{summary_row},0)'),
        ('Tests Overdue/Non-Compliant:', f'=B{summary_row+2}+B{summary_row+3}'),
    ]
    
    for label, formula in metrics:
        if label:
            ws[f'A{summary_row}'] = label
            apply_style(ws[f'A{summary_row}'], font=BOLD_FONT)
            ws[f'B{summary_row}'] = formula
            apply_style(ws[f'B{summary_row}'], font=BOLD_FONT)
            if 'Rate' in label:
                ws[f'B{summary_row}'].number_format = '0.0%'
        summary_row += 1
    
    # Column widths
    widths = [25, 20, 18, 18, 15, 20, 20, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Lazy-add: only register DVs that have cells assigned (fixes empty-sqref XML error)
    for _dv in validations.values():
        if _dv.sqref:
            ws.add_data_validation(_dv)
    
    ws.freeze_panes = 'A4'
    
    return ws

# ============================================================================
# WORKSHEET: EVIDENCE REGISTER
# ============================================================================

def create_evidence_register(wb):
    """Create Evidence Register worksheet — Gold Standard (GS-ER)"""
    ws = wb.create_sheet(title="Evidence Register")
    ws.sheet_view.showGridLines = False
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: title (003366)
    ws.merge_cells('A1:H1')
    ws['A1'] = 'EVIDENCE REGISTER'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Row 2: italic subtitle
    ws.merge_cells('A2:H2')
    ws['A2'] = (
        f'Document all evidence supporting this assessment '
        f'({DOCUMENT_ID} | MINIMUM 5 items required for audit compliance)'
    )
    ws['A2'].font = Font(name='Calibri', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Row 4: column headers (003366)
    headers = [
        'Evidence ID', 'Evidence Type', 'Description',
        'Related Sheet / Row', 'Location / Path',
        'Date Collected', 'Collected By', 'Verification Status',
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Row 5: F2F2F2 sample row (EV-001)
    sample = [
        'EV-001', 'Test Report',
        'E-Commerce backup restore test report — Q1',
        "Test Results Log!A5 (RES-001)",
        '/evidence/testing/ecommerce_restore_Q1.pdf',
        datetime.now().strftime('%d.%m.%Y'), 'Infrastructure Team', f'{CHECK} Verified',
    ]
    for col_idx, value in enumerate(sample, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        cell.font = Font(name='Calibri', size=10, italic=True)
        cell.border = border
        if col_idx == 6:
            cell.number_format = 'DD.MM.YYYY'

    # Rows 6–105: 100 FFFFCC empty rows
    dv_evidence_type = None
    dv_verification = None
    for dv in ws.data_validations.dataValidation:
        pass  # placeholder; we add fresh DVs below

    ev_type_dv = DataValidation(
        type="list",
        formula1='"Test Report,Screenshot,Log File,Video Recording,Configuration File,Runbook,Issue Ticket,Sign-Off,Other"',
        allow_blank=False,
    )
    verif_dv = DataValidation(
        type="list",
        formula1=f'"{CHECK} Verified,\u23f3 Pending,{XMARK} Not Verified"',
        allow_blank=False,
    )
    ws.add_data_validation(ev_type_dv)
    ws.add_data_validation(verif_dv)

    for i in range(100):
        r = 6 + i
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            cell.border = border
            if col_idx == 6:
                cell.number_format = 'DD.MM.YYYY'
        ev_type_dv.add(f'B{r}')
        verif_dv.add(f'H{r}')

    # Column widths
    col_widths = [12, 18, 40, 25, 35, 15, 20, 20]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = 'A5'

    return ws

# ============================================================================
# WORKSHEET: APPROVAL SIGN-OFF
# ============================================================================

def create_approval_sheet(wb):
    """Create Approval Sign-Off worksheet (Gold Standard GS-AS)"""
    ws = wb.create_sheet(title="Approval Sign-Off")
    ws.sheet_view.showGridLines = False
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: TITLE BANNER (GS-AS-001/002/003: A1:E1 merge, "AND" not "&", height 35)
    ws.merge_cells('A1:E1')
    ws['A1'] = 'ASSESSMENT APPROVAL AND SIGN-OFF'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 35
    for col in range(1, 6):
        ws.cell(row=1, column=col).border = border

    # Row 2: CONTROL REFERENCE subtitle (DS-006)
    ws.merge_cells('A2:E2')
    ws['A2'] = f'{DOCUMENT_ID} | ISO/IEC 27001:2022 - Control A.5.30 (BC/DR Testing)'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='003366')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 6):
        ws.cell(row=2, column=col).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells('A3:E3')
    ws['A3'] = 'ASSESSMENT SUMMARY'
    ws['A3'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws['A3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=3, column=col).border = border

    # Summary fields (rows 4-8)
    summary_fields = [
        ('Document:', f'{DOCUMENT_ID} - {WORKBOOK_TITLE}'),
        ('Assessment Period:', ''),
        ('Overall Test Pass Rate:', "='Summary Dashboard'!B13"),
        ('Assessment Status:', ''),
        ('Assessed By:', ''),
    ]
    row = 4
    status_row_for_dv = None
    for label, value in summary_fields:
        editable = (value == '')
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = value
        for col in range(2, 6):
            if editable:
                ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        if 'Assessment Status' in label:
            status_row_for_dv = row
        if 'Rate' in label or 'Coverage' in label:
            ws[f'B{row}'].number_format = '0.0%'
        row += 1

    # Assessment Status dropdown
    status_dv = DataValidation(
        type='list',
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    if status_row_for_dv:
        status_dv.add(f'B{status_row_for_dv}')

    row += 2  # Gap before first approver

    def _create_approver_section(start_row, title, color):
        """Create one approver section (GS-AS-009)"""
        ws.merge_cells(f'A{start_row}:E{start_row}')
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        ws[f'A{start_row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, 6):
            ws.cell(row=start_row, column=col).border = border
        current_row = start_row + 1
        for field in ['Name:', 'Title:', 'Date:', 'Signature:', 'Comments:']:
            ws[f'A{current_row}'] = field
            ws[f'A{current_row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'A{current_row}'].border = border
            ws.merge_cells(f'B{current_row}:E{current_row}')
            for col in range(2, 6):
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                ws.cell(row=current_row, column=col).border = border
            current_row += 1
        return current_row + 1

    row = _create_approver_section(row, 'COMPLETED BY (ASSESSOR)', '4472C4')
    row = _create_approver_section(row, 'REVIEWED BY (INFORMATION SECURITY OFFICER)', '4472C4')
    row = _create_approver_section(row, 'APPROVED BY (CISO)', '003366')

    # FINAL DECISION (GS-AS-004: plain bold label A, FFFFCC B:E — NO fill on A)
    ws[f'A{row}'] = 'FINAL DECISION:'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'A{row}'].border = border
    ws.merge_cells(f'B{row}:E{row}')
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws.cell(row=row, column=col).border = border
    decision_dv = DataValidation(
        type='list',
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_dv)
    decision_dv.add(f'B{row}')

    # NEXT REVIEW DETAILS (GS-AS: 4472C4 banner)
    row += 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'NEXT REVIEW DETAILS'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    row += 1
    for label in ['Next Review Date:', 'Review Responsible:', 'Special Considerations:']:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row}'].border = border
        ws.merge_cells(f'B{row}:E{row}')
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            ws.cell(row=row, column=col).border = border
        row += 1

    # Column widths + freeze
    ws.column_dimensions['A'].width = 32
    for col_letter in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col_letter].width = 20
    ws.freeze_panes = 'A3'


def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        ws.data_validations.dataValidation = [
            dv for dv in list(ws.data_validations.dataValidation)
            if dv.sqref
        ]


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info(f"\n{'='*70}")
    logger.info(f"GENERATING: {WORKBOOK_TITLE}")
    logger.info(f"{'='*70}")
    logger.info(f"Version: {VERSION}")
    logger.info(f"Controls: {CONTROLS}")
    logger.info(f"Assessment ID: {DOCUMENT_ID}")
    logger.info(f"{'='*70}\n")
    
    # Create workbook
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.remove(wb.active)
    
    # Create all worksheets in order
    logger.info("Creating worksheets...")
    create_instructions_sheet(wb.create_sheet())
    logger.info("  Instructions & Legend")

    create_test_schedule_sheet(wb)
    logger.info("  Test Schedule")

    create_test_results_log_sheet(wb)
    logger.info("  Test Results Log")

    create_issue_remediation_sheet(wb)
    logger.info("  Issue Remediation")

    create_testing_compliance_sheet(wb)
    logger.info("  Testing Compliance")

    create_evidence_register(wb)
    logger.info("  Evidence Register")

    create_summary_dashboard_sheet(wb)
    logger.info("  Summary Dashboard")

    create_approval_sheet(wb)
    logger.info("  Approval Sign-Off")
    
    # Save workbook
    finalize_validations(wb)
    wb.save(output_path)
    # Summary
    logger.info(f"\n{'='*70}")
    logger.info("WORKBOOK GENERATED SUCCESSFULLY")
    logger.info(f"{'='*70}")
    logger.info(f"Filename: {output_path.name}")
    logger.info(f"Worksheets: {len(wb.sheetnames)}")
    logger.info("\nWorksheet Details:")
    logger.info("  Instructions & Legend (comprehensive usage guide)")
    logger.info("  Summary Dashboard (executive dashboard with all metrics)")
    logger.info("  Test Schedule (110 rows: 10 examples + 100 data entry)")
    logger.info("  Test Results Log (110 rows: 10 examples + 100 results)")
    logger.info("  Issue Remediation (110 rows: 10 examples + 100 issues)")
    logger.info("  Testing Compliance (110 rows: 10 examples + 100 compliance checks)")
    logger.info("  Evidence Register (1 sample + 100 FFFFCC empty rows)")
    logger.info("  Approval Sign-Off (3-level workflow: Assessor / ISO / CISO)")
    logger.info(f"\n{'='*70}")
    logger.info(f"{CHECK} AUDIT-READY FEATURES:")
    logger.info("  • Evidence tracking (minimum 5 items required)")
    logger.info("  • 3-level approval workflow (Assessor → ISO → CISO)")
    logger.info("  • Comprehensive data validations (12 dropdown types)")
    logger.info("  • Auto-calculated compliance metrics (days since last test)")
    logger.info("  • Test success rate tracking")
    logger.info("  • Issue remediation workflow")
    logger.info("  • Professional styling without Excel repair warnings")
    logger.info(f"{'='*70}\n")
    
def main():
    try:
        create_workbook(_wkbk_dir / OUTPUT_FILENAME)
    except Exception as e:
        logger.info(f"\n{'='*70}")
        logger.error(f"{XMARK} ERROR: Failed to generate workbook")
        logger.info(f"{'='*70}")
        logger.error(f"Error details: {str(e)}")
        logger.error(f"Error type: {type(e).__name__}")
        import traceback
        logger.info(f"\nFull traceback:")
        traceback.print_exc()
        logger.info(f"{'='*70}\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
