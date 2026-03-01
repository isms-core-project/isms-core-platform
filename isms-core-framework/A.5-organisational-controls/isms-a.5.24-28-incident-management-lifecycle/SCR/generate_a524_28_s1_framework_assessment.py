#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.24-28.S1 - Incident Management Framework Assessment Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.24: Incident Management Planning & Preparation
Assessment Domain 1 of 5: Framework & Governance

--------------------------------------------------------------------------------
PURPOSE & DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
the incident management framework and governance structure, focusing on the
planning and preparation phase (A.5.24) of the incident management lifecycle.

**Assessment Scope:**
- Governance framework (policy, procedures, authority)
- CSIRT/SOC organisational structure and staffing
- Role definitions, responsibilities, and RACI clarity
- Training and competency programs
- Tools and technology capabilities
- Integration with monitoring, logging, BC/DR, and other controls

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and methodology
2. Governance Assessment - 25 questions on policy, procedures, compliance
3. Organisational Structure - 30 questions on CSIRT model, roles, staffing
4. Training & Competency - 25 questions on programs, exercises, certifications
5. Tools & Technology - 30 questions on ticketing, SIEM, forensics, automation
6. Integration Assessment - 25 questions on cross-control integration
7. Gap Analysis - Gap prioritization and remediation planning
8. Evidence Register - Audit evidence tracking (100 items capacity)
9. Dashboard - Executive summary with maturity scoring
10. Approval Sign-Off - Multi-stakeholder approval workflow

**Key Features:**
- 117 assessment questions across 5 domains
- Data validation with dropdown lists
- Conditional formatting for gap identification
- Automated maturity scoring (Levels 1-5)
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This is Domain 1 of 5 assessment domains for A.5.24-28 Incident Management:
- S1: Framework & Governance (this assessment)
- S2: Detection & Classification (A.5.25 focus)
- S3: Response Capabilities (A.5.26 focus)
- S4: Forensic Evidence Management (A.5.28 focus)
- S5: Learning & Improvement (A.5.27 focus)

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a524_28_s1_framework_assessment.py

Output:
    File: ISMS-IMP-A.5.24-28.S1_Framework_Assessment_YYYYMMDD.xlsx
    Location: Current directory

Post-Generation Steps:
    1. Review and customize assessment questions for your organisation
    2. Complete document information in Instructions & Legend
    3. Conduct stakeholder interviews (CSIRT, Legal, HR, IT Ops)
    4. Fill yellow cells in each assessment sheet (Governance through Integration)
    5. Collect and link audit evidence (policy docs, org charts, training records)
    6. Review Gap Analysis and prioritize remediation
    7. Complete Dashboard review with management
    8. Obtain stakeholder approvals (CSIRT Manager, CISO)

Estimated Completion Time:
    - Information gathering: 3-4 hours
    - Assessment completion: 3-4 hours
    - Evidence collection: 1-2 hours
    - Quality review: 1-2 hours
    Total: 8-12 hours (depending on organisation size/complexity)

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.5.24-28
Assessment Domain:    1 of 5 (Framework & Governance - A.5.24 Focus)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              Internal Use

Related Documents:
    - ISMS-POL-A.5.24-28: Incident Management Lifecycle Policy
    - ISMS-REF-A.5.24-28: Incident Response Reference Guide
    - ISMS-IMP-A.5.24-28.S2: Detection & Classification Assessment
    - ISMS-IMP-A.5.24-28.S3: Response Capabilities Assessment
    - ISMS-IMP-A.5.24-28.S4: Forensic Evidence Assessment
    - ISMS-IMP-A.5.24-28.S5: Learning & Improvement Assessment

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - 2026-01-30
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.24-28.S1 specification
    - 117 assessment questions across 5 domains
    - Automated maturity scoring and gap analysis
    - Integration with evidence tracking

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Customization Required:**
This script generates a template assessment. Customize the following:
- Question relevance to your organisational context
- Dropdown options to match your CSIRT model and tools
- Evidence types to align with your audit requirements
- Maturity scoring thresholds based on your risk appetite

**Collaboration Essential:**
This assessment CANNOT be completed in isolation. Required stakeholders:
- CSIRT/SOC Manager (primary assessor)
- CISO (governance, budget)
- HR (staffing, training records)
- IT Operations (tools, integration)
- Legal/Compliance (regulatory requirements)

**Assessment Frequency:**
- Annual review (minimum)
- After major organisational changes (merger, restructuring)
- After major incidents (lessons learned)
- After audit findings

**Data Sensitivity:**
Assessment workbooks contain sensitive information:
- Organisational structure and staffing details
- Tool inventory and capabilities
- Security gaps and vulnerabilities
- Training records and competency data

Handle in accordance with your organisation's data classification policies.

================================================================================
"""

from datetime import datetime
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging
import sys

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.24-28.S1"
WORKBOOK_NAME = "Incident Management Framework"
CONTROL_ID = "A.5.24-28"
CONTROL_NAME = "Incident Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching specification."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure per markdown specification
    sheets = [
        "Instructions & Legend",
        "Governance Assessment",
        "Organisational Structure",
        "Training & Competency",
        "Tools & Technology",
        "Integration Assessment",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "calculated_cell": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
        "gap_critical": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "gap_high": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "gap_medium": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "gap_low": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
    }
    return styles


# ============================================================================
# SECTION 2: INSTRUCTIONS & LEGEND SHEET
# ============================================================================



_STYLES = setup_styles()
def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete Governance Assessment (25 questions on policy, procedures, authority).', '2. Complete Organisational Structure assessment (30 questions on CSIRT/SOC model).', '3. Complete Training & Competency assessment (25 questions).', '4. Complete Tools & Technology assessment (30 questions).', '5. Complete Integration Assessment (25 questions on cross-control integration).', '6. Complete Gap Analysis with prioritised remediation actions.', '7. Link audit evidence in the Evidence Register.', '8. Review Summary Dashboard for maturity scoring.', '9. Obtain stakeholder approvals in Approval Sign-Off.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_governance_assessment(ws, styles):
    """Create the Governance Assessment sheet (25 questions)."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "GOVERNANCE ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A2"] = f"{CONTROL_REF} | Policy, Procedures, Authority, and Regulatory Compliance (25 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")

    # Column headers
    row = 4
    headers = ["Question ID", "Section", "Question", "Answer", "Evidence Reference", "Comments", "Gap Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Questions (25 total across 6 sections)
    questions = [
        # Section A: Policy Documentation (Q1-Q5)
        ("Q1", "Policy Documentation", "Does [Organisation] have a documented incident management policy?", "Yes/No/In Draft"),
        ("Q2", "Policy Documentation", "When was the current incident management policy approved? (DD.MM.YYYY)", "Date"),
        ("Q3", "Policy Documentation", "Who is the designated owner of the incident management policy? (Name, Title)", "Text"),
        ("Q4", "Policy Documentation", "How often is the incident management policy reviewed?", "Annually/Bi-Annually/Quarterly/Ad-Hoc/Not Defined"),
        ("Q5", "Policy Documentation", "When was the policy last reviewed? (DD.MM.YYYY)", "Date"),
        
        # Section B: Procedure Documentation (Q6-Q10)
        ("Q6", "Procedure Documentation", "Are incident response procedures documented?", "Yes - Comprehensive/Yes - Partial/No/In Development"),
        ("Q7", "Procedure Documentation", "In what format are incident response procedures documented? (Select all: Written/Flowcharts/Automated/Scenarios/Other)", "Text"),
        ("Q8", "Procedure Documentation", "Where are incident response procedures stored and how are they accessed?", "Text"),
        ("Q9", "Procedure Documentation", "Do procedures include up-to-date contact information (CSIRT, management, legal, external partners)?", "Yes - Current/Yes - Outdated/No"),
        ("Q10", "Procedure Documentation", "When were incident response procedures last updated? (DD.MM.YYYY)", "Date"),
        
        # Section C: Authority & Escalation (Q11-Q15)
        ("Q11", "Authority & Escalation", "Who has authority to declare a security incident? (Role/title)", "Text"),
        ("Q12", "Authority & Escalation", "Is there a documented escalation matrix defining when and to whom incidents are escalated?", "Yes - Comprehensive/Yes - Basic/No"),
        ("Q13", "Authority & Escalation", "What criteria trigger executive management notification?", "Text"),
        ("Q14", "Authority & Escalation", "What criteria trigger Board of Directors notification?", "Text"),
        ("Q15", "Authority & Escalation", "Are procedures defined for external notifications (regulators, law enforcement, customers, media)?", "Yes - Documented/Partially Documented/No"),
        
        # Section D: Regulatory Compliance (Q16-Q20)
        ("Q16", "Regulatory Compliance", "Have applicable regulatory incident notification requirements been identified?", "Yes - Documented/Partially/No"),
        ("Q17", "Regulatory Compliance", "(If GDPR applies) Is GDPR 72-hour breach notification procedure documented?", "Yes/No/N/A"),
        ("Q18", "Regulatory Compliance", "(If Swiss nDSG applies) Is Swiss nDSG breach notification procedure documented?", "Yes/No/N/A"),
        ("Q19", "Regulatory Compliance", "(If PCI DSS v4.0.1 applies) Is PCI DSS v4.0.1 incident reporting procedure documented?", "Yes/No/N/A"),
        ("Q20", "Regulatory Compliance", "Are sector-specific incident reporting requirements documented (FINMA, DORA, NIS2, HIPAA, etc.)?", "Yes/No/N/A"),
        
        # Section E: Exception Management (Q21-Q22)
        ("Q21", "Exception Management", "Is there a process for exceptions to incident management requirements?", "Yes/No"),
        ("Q22", "Exception Management", "Are current exceptions documented and approved?", "Yes/No/N/A (no exceptions)"),
        
        # Section F: Policy Governance (Q23-Q25)
        ("Q23", "Policy Governance", "How is the incident management policy communicated to staff? (Select all: Repository/Onboarding/Training/Other)", "Text"),
        ("Q24", "Policy Governance", "Can relevant staff easily access the incident management policy?", "Yes - Widely Accessible/Yes - Restricted Access/No"),
        ("Q25", "Policy Governance", "When is the next scheduled policy review? (DD.MM.YYYY)", "Date"),
    ]

    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]  # Answer cell (yellow)
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]  # Evidence (yellow)
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]  # Comments (yellow)
        
        # Gap formula in column G
        ws[f"G{row}"] = f'=IF(D{row}="","",IF(OR(D{row}="No", D{row}="Partial", D{row}="Informal"), "Yes", "No"))'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12


    # Answer column DV (D column) — provides dropdown for gap-relevant answers
    from openpyxl.worksheet.datavalidation import DataValidation as _DV
    _dv_answer = _DV(type="list", formula1='"Yes,No,Partial,Informal,In Development,In Draft,N/A"', allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(_dv_answer)
    for _r in range(5, row):
        _dv_answer.add(f"D{_r}")

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 4: ORGANIZATIONAL STRUCTURE SHEET
# ============================================================================

def create_organisational_structure(ws, styles):
    """Create the Organisational Structure sheet (30 questions)."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "ORGANISATIONAL STRUCTURE"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A2"] = f"{CONTROL_REF} | CSIRT/SOC Model, Roles, Staffing, and RACI (30 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")

    # Column headers
    row = 4
    headers = ["Question ID", "Section", "Question", "Answer", "Evidence Reference", "Comments", "Gap Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Questions (30 total across 5 sections)
    questions = [
        # Section A: CSIRT/SOC Model (Q26-Q30)
        ("Q26", "CSIRT/SOC Model", "What incident response organisational model does [Organisation] use?", "Dedicated CSIRT/Virtual CSIRT/Hybrid/Outsourced SOC/Coordinated"),
        ("Q27", "CSIRT/SOC Model", "When was the CSIRT/SOC team established? (MM.YYYY)", "Date"),
        ("Q28", "CSIRT/SOC Model", "What coverage does the CSIRT/SOC provide?", "24/7/365/Business Hours/Extended Hours/On-Call Rotation"),
        ("Q29", "CSIRT/SOC Model", "How many full-time equivalent (FTE) staff are dedicated to incident response?", "Number"),
        ("Q30", "CSIRT/SOC Model", "To whom does the CSIRT/SOC team report? (Title/Department)", "Text"),
        
        # Section B: Role Definitions (Q31-Q45 - 15 roles, showing first 5 here for brevity)
        ("Q31", "Role Definitions", "Incident Manager - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q32", "Role Definitions", "SOC Analyst Tier 1 - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q33", "Role Definitions", "SOC Analyst Tier 2 - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q34", "Role Definitions", "Forensic Specialist - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q35", "Role Definitions", "Malware Analyst - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q36", "Role Definitions", "Threat Intelligence Analyst - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q37", "Role Definitions", "Communication Coordinator - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q38", "Role Definitions", "Legal Representative - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q39", "Role Definitions", "HR Representative - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q40", "Role Definitions", "IT Operations Liaison - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q41", "Role Definitions", "Business Unit Liaison - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q42", "Role Definitions", "Third-Party Coordinator - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q43", "Role Definitions", "Executive Sponsor - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q44", "Role Definitions", "On-Call Engineer - Does role exist? (Yes/No), Assigned to:", "Text"),
        ("Q45", "Role Definitions", "Other Roles - List any additional incident response roles:", "Text"),
        
        # Section C: RACI Matrix (Q46-Q48)
        ("Q46", "RACI Matrix", "Is there a documented RACI matrix for incident management roles?", "Yes - Comprehensive/Yes - Basic/No"),
        ("Q47", "RACI Matrix", "Overall, are roles and responsibilities clear and unambiguous?", "Yes - Very Clear/Mostly Clear/Some Ambiguity/Significant Confusion"),
        ("Q48", "RACI Matrix", "When was the RACI matrix last reviewed/updated? (DD.MM.YYYY)", "Date"),
        
        # Section D: Staffing Adequacy (Q49-Q52)
        ("Q49", "Staffing Adequacy", "Is current staffing adequate to support stated coverage?", "Yes/No/Barely Adequate"),
        ("Q50", "Staffing Adequacy", "What is the CSIRT/SOC annual turnover rate? (%)", "Percentage"),
        ("Q51", "Staffing Adequacy", "Is there succession planning for key incident response roles?", "Yes/Informal/No"),
        ("Q52", "Staffing Adequacy", "Have skill gaps in the incident response team been identified?", "Yes - Documented/Informally Identified/No"),
        
        # Section E: Management Visibility (Q53-Q55)
        ("Q53", "Management Visibility", "How frequently does CSIRT report to management (CISO, CIO, or Executive)?", "Weekly/Monthly/Quarterly/Ad-Hoc/Never"),
        ("Q54", "Management Visibility", "Are incident management performance metrics defined and tracked?", "Yes - Comprehensive/Yes - Basic/No"),
        ("Q55", "Management Visibility", "Is there a protocol for briefing the Board of Directors on material incidents?", "Yes/Informal/No/N/A"),
    ]

    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(D{row}="","",IF(OR(D{row}="No", D{row}="Vacant", D{row}="Barely Adequate"), "Yes", "No"))'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12


    # Answer column DV (D column) — provides dropdown for gap-relevant answers
    from openpyxl.worksheet.datavalidation import DataValidation as _DV
    _dv_answer = _DV(type="list", formula1='"Fully Staffed,Partially Staffed,Vacant,Barely Adequate,Yes,No,Informal,N/A"', allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(_dv_answer)
    for _r in range(5, row):
        _dv_answer.add(f"D{_r}")

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 5: TRAINING & COMPETENCY SHEET
# ============================================================================

def create_training_competency(ws, styles):
    """Create the Training & Competency sheet (25 questions)."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "TRAINING & COMPETENCY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A2"] = f"{CONTROL_REF} | Training Programs, Exercises, and Certifications (25 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")

    # Column headers
    row = 4
    headers = ["Question ID", "Section", "Question", "Answer", "Evidence Reference", "Comments", "Gap Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Questions (25 total across 5 sections)
    questions = [
        # Section A: Training Program (Q56-Q60)
        ("Q56", "Training Program", "Does [Organisation] have a documented incident response training program?", "Yes - Comprehensive/Yes - Basic/No"),
        ("Q57", "Training Program", "Do all staff receive annual security awareness training including incident reporting?", "Yes/Partial/No"),
        ("Q58", "Training Program", "What percentage of staff completed incident reporting awareness training in the last 12 months? (%)", "Percentage"),
        ("Q59", "Training Program", "Do CSIRT members receive specialized incident response training?", "Yes - Regular/Yes - Occasional/No"),
        ("Q60", "Training Program", "Is there a dedicated budget for incident response training (courses, certifications, conferences)?", "Yes - Adequate/Yes - Limited/No"),
        
        # Section B: Tabletop Exercises (Q61-Q65)
        ("Q61", "Tabletop Exercises", "How often does [Organisation] conduct incident response tabletop exercises?", "Quarterly/Semi-Annually/Annually/Bi-Annually/Never"),
        ("Q62", "Tabletop Exercises", "When was the most recent tabletop exercise conducted? (DD.MM.YYYY)", "Date"),
        ("Q63", "Tabletop Exercises", "Do tabletop exercises cover a variety of incident types?", "Yes - Diverse Scenarios/Limited Variety/Single Scenario"),
        ("Q64", "Tabletop Exercises", "Who typically participates in tabletop exercises? (Select all: CSIRT/IT Ops/Legal/HR/Management/Executive/Board/External)", "Text"),
        ("Q65", "Tabletop Exercises", "Are lessons learned from tabletop exercises documented and acted upon?", "Yes - Systematically/Yes - Informally/No"),
        
        # Section C: Simulation & Testing (Q66-Q68)
        ("Q66", "Simulation & Testing", "Does [Organisation] conduct full-scale incident response exercises (beyond tabletop)?", "Yes - Regularly/Yes - Occasionally/No"),
        ("Q67", "Simulation & Testing", "Does [Organisation] use breach simulation or attack simulation tools?", "Yes/No"),
        ("Q68", "Simulation & Testing", "Are drills conducted specifically for regulatory breach notification (GDPR 72-hour, PCI DSS v4.0.1, etc.)?", "Yes/No"),
        
        # Section D: Competency & Certification (Q69-Q72)
        ("Q69", "Competency & Certification", "Are competency requirements defined for incident response roles?", "Yes/No"),
        ("Q70", "Competency & Certification", "Does [Organisation] track security certifications for CSIRT members?", "Yes/Informal/No"),
        ("Q71", "Competency & Certification", "Are certifications current (not expired)?", "Yes - All Current/Some Expired/Not Tracked"),
        ("Q72", "Competency & Certification", "Is there a documented onboarding process for new CSIRT members?", "Yes - Formal Program/Informal/No"),
        
        # Section E: Training Effectiveness (Q73-Q75)
        ("Q73", "Training Effectiveness", "Is training effectiveness measured (beyond completion tracking)?", "Yes/No"),
        ("Q74", "Training Effectiveness", "Is feedback collected from training participants?", "Yes - Systematically/Occasionally/No"),
        ("Q75", "Training Effectiveness", "Have training gaps been identified through exercises or actual incidents?", "Yes - Documented/Informally Identified/No"),
        
        # Additional questions to reach 25
        ("Q76", "Training Records", "Are training records maintained and accessible for audit purposes?", "Yes/Partial/No"),
        ("Q77", "Training Records", "How long are training records retained? (Years)", "Number"),
        ("Q78", "External Training", "Have CSIRT members attended external training in the last 12 months?", "Yes/No"),
        ("Q79", "Certification Goals", "Are certification goals set for CSIRT members?", "Yes/No"),
        ("Q80", "Training Budget", "What percentage of security budget is allocated to incident response training? (%)", "Percentage"),
    ]

    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(D{row}="","",IF(OR(D{row}="No", D{row}<95, D{row}="Never"), "Yes", "No"))'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12


    # Answer column DV (D column) — provides dropdown for gap-relevant answers
    from openpyxl.worksheet.datavalidation import DataValidation as _DV
    _dv_answer = _DV(type="list", formula1='"Yes,No,In Progress,Never,N/A"', allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(_dv_answer)
    for _r in range(5, row):
        _dv_answer.add(f"D{_r}")

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 6: TOOLS & TECHNOLOGY SHEET
# ============================================================================

def create_tools_technology(ws, styles):
    """Create the Tools & Technology sheet (30 questions)."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "TOOLS & TECHNOLOGY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A2"] = f"{CONTROL_REF} | Ticketing, SIEM, Forensics, Communication, and Automation (30 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")

    # Column headers
    row = 4
    headers = ["Question ID", "Section", "Question", "Answer", "Evidence Reference", "Comments", "Gap Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Questions (30 total across 7 sections)
    questions = [
        # Section A: Ticketing (Q76-Q79)
        ("Q81", "Ticketing", "Does [Organisation] have a dedicated incident ticketing/tracking system?", "Yes - Dedicated IR/Yes - General IT/No"),
        ("Q82", "Ticketing", "What features does the incident ticketing system provide? (Select all: Creation/Severity/Timeline/Evidence/Tasks/Collaboration/Escalation/Integration/Reporting/Playbooks)", "Text"),
        ("Q83", "Ticketing", "How long are incident tickets retained? (Years)", "1 year/3 years/5 years/7+ years/Indefinitely/Not Defined"),
        ("Q84", "Ticketing", "Are access controls implemented for incident tickets (need-to-know basis)?", "Yes - Strict/Yes - Moderate/No - Open"),
        
        # Section B: SIEM (Q80-Q83)
        ("Q85", "SIEM", "Does [Organisation] have a Security Information and Event Management (SIEM) system?", "Yes - Fully Operational/Yes - Partial/No"),
        ("Q86", "SIEM", "How many log sources feed into the SIEM? (Number)", "Number"),
        ("Q87", "SIEM", "Is the SIEM integrated with the incident ticketing system (automatic ticket creation)?", "Yes - Fully Automated/Yes - Semi-Automated/Manual/N/A"),
        ("Q88", "SIEM", "Does the SIEM/SOAR platform support automated playbooks for common incident types?", "Yes - Extensive/Yes - Limited/No/N/A"),
        
        # Section C: Forensics (Q84-Q87)
        ("Q89", "Forensics", "What digital forensic tools are available? (Select all: Disk imaging/Memory acquisition/Memory analysis/Network forensics/Log analysis/Malware analysis/Mobile/Cloud/Commercial suites/None)", "Text"),
        ("Q90", "Forensics", "Are CSIRT members trained in the use of forensic tools?", "Yes - All Members/Yes - Specialized Roles/Limited/No"),
        ("Q91", "Forensics", "Does [Organisation] have a dedicated forensic workstation or lab?", "Yes - Dedicated Lab/Yes - Virtual/No"),
        ("Q92", "Forensics", "Is there secure storage for forensic evidence?", "Yes - Dedicated/Yes - General/No"),
        
        # Section D: Communication (Q88-Q90)
        ("Q93", "Communication", "What tools are used for internal incident response communication? (Select all: Email/Phone/IM/Video/War Room/IR Platform/Encrypted)", "Text"),
        ("Q94", "Communication", "Is there a secure channel for external communications (legal, law enforcement, regulators)?", "Yes/No"),
        ("Q95", "Communication", "How are on-call incident responders alerted?", "PagerDuty/VictorOps/Phone/SMS/Email/No System"),
        
        # Section E: Threat Intel (Q91-Q93)
        ("Q96", "Threat Intel", "Does [Organisation] subscribe to threat intelligence feeds?", "Yes - Multiple/Yes - Single/No"),
        ("Q97", "Threat Intel", "Is threat intelligence integrated into SIEM for alerting?", "Yes - Automated/Yes - Manual/No/N/A"),
        ("Q98", "Threat Intel", "Is there a dedicated threat intelligence analyst role?", "Yes - Dedicated/Yes - Part-Time/No"),
        
        # Section F: Automation (Q94-Q96)
        ("Q99", "Automation", "What level of incident response automation exists?", "High (SOAR)/Moderate (Some Playbooks)/Low (Mostly Manual)/None"),
        ("Q100", "Automation", "How many documented/automated playbooks exist? (Number)", "Number"),
        ("Q101", "Automation", "Are playbooks regularly reviewed and updated?", "Yes - Annually/Yes - After Incidents/No/N/A"),
        
        # Section G: Tool Gaps (Q97-Q99)
        ("Q102", "Tool Gaps", "Are there critical tool gaps preventing effective incident response?", "Yes - Multiple/Yes - Some/No"),
        ("Q103", "Tool Gaps", "Are tool gaps being addressed (budget approved, procurement in progress)?", "Yes - Funded/Proposed/No"),
        ("Q104", "Tool Gaps", "Are there challenges integrating incident response tools?", "Yes - Significant/Yes - Moderate/No/N/A"),
        
        # Additional questions
        ("Q105", "Tool Inventory", "Is there a documented inventory of all incident response tools?", "Yes/No"),
        ("Q106", "Tool Licensing", "Are all incident response tools properly licensed?", "Yes/Some Unlicensed/Not Tracked"),
        ("Q107", "Tool Maintenance", "Are incident response tools regularly updated and maintained?", "Yes/Partial/No"),
        ("Q108", "Tool Training", "Do CSIRT members receive training on all critical tools?", "Yes/Partial/No"),
        ("Q109", "Tool Documentation", "Is tool documentation maintained and accessible?", "Yes/Partial/No"),
        ("Q110", "Tool Testing", "Are incident response tools tested regularly?", "Yes/Occasional/No"),
    ]

    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(D{row}="","",IF(OR(D{row}="No", D{row}="None", D{row}="N/A"), "Yes", "No"))'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12


    # Answer column DV (D column) — provides dropdown for gap-relevant answers
    from openpyxl.worksheet.datavalidation import DataValidation as _DV
    _dv_answer = _DV(type="list", formula1='"Yes,No,None,Partial,In-Progress,N/A"', allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(_dv_answer)
    for _r in range(5, row):
        _dv_answer.add(f"D{_r}")

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 7: INTEGRATION ASSESSMENT SHEET
# ============================================================================

def create_integration_assessment(ws, styles):
    """Create the Integration Assessment sheet (25 questions)."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "INTEGRATION ASSESSMENT"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A2"] = f"{CONTROL_REF} | Integration with Logging, Monitoring, BC/DR, and Other Controls (25 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")

    # Column headers
    row = 4
    headers = ["Question ID", "Section", "Question", "Answer", "Evidence Reference", "Comments", "Gap Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Questions (25 total across 6 sections)
    questions = [
        # Section A: Logging Integration (Q100-Q102)
        ("Q111", "Logging Integration", "Is incident management integrated with logging infrastructure (A.8.15)?", "Yes - Fully Integrated/Partially Integrated/No"),
        ("Q112", "Logging Integration", "Are logs readily available to CSIRT during incidents?", "Yes - Immediate Access/Yes - Delayed Access/No"),
        ("Q113", "Logging Integration", "Is log retention adequate for incident investigation?", "Yes/Insufficient for Some Systems/No"),
        
        # Section B: Monitoring Integration (Q103-Q105)
        ("Q114", "Monitoring Integration", "Is incident management integrated with security monitoring (A.8.16)?", "Yes - Fully Integrated/Partially Integrated/No"),
        ("Q115", "Monitoring Integration", "Is there a defined workflow from security alert to incident declaration?", "Yes - Documented/Informal/No"),
        ("Q116", "Monitoring Integration", "Is the false positive rate from monitoring tracked?", "Yes/No"),
        
        # Section C: User Reporting Integration (Q106-Q108)
        ("Q117", "User Reporting", "Is there a mechanism for users to report security events (A.6.8 integration)?", "Yes - Dedicated/Yes - General IT Helpdesk/No"),
        ("Q118", "User Reporting", "Are user-reported events integrated into incident ticketing system?", "Yes - Automatically/Yes - Manually/No"),
        ("Q119", "User Reporting", "Do users know how to report security incidents?", "Yes - Well Communicated/Somewhat/No"),
        
        # Section D: BC/DR Integration (Q109-Q111)
        ("Q120", "BC/DR Integration", "Is incident management integrated with Business Continuity / Disaster Recovery (A.5.29-30)?", "Yes - Documented/Informal/No"),
        ("Q121", "BC/DR Integration", "Are criteria defined for when an incident triggers BC/DR activation?", "Yes/No"),
        ("Q122", "BC/DR Integration", "Is CSIRT's role defined in BC/DR plans?", "Yes - Documented/Informal/No"),
        
        # Section E: Third-Party Coordination (Q112-Q114)
        ("Q123", "Third-Party Coordination", "Are procedures defined for incidents involving third parties (suppliers, MSSPs)?", "Yes - Documented/Informal/No"),
        ("Q124", "Third-Party Coordination", "(If MSSP/outsourced SOC) Are SLAs defined for incident detection and escalation?", "Yes/No/N/A"),
        ("Q125", "Third-Party Coordination", "Do suppliers know how to report security incidents to [Organisation]?", "Yes - Defined in Contracts/Informal/No"),
        
        # Section F: Legal & Regulatory (Q115-Q117)
        ("Q126", "Legal & Regulatory", "Is Legal/Compliance involvement in incident management clearly defined?", "Yes - Documented/Informal/No"),
        ("Q127", "Legal & Regulatory", "(If GDPR applicable) Is the Data Protection Officer (DPO) integrated into incident management?", "Yes/No/N/A"),
        ("Q128", "Legal & Regulatory", "Are procedures defined for coordinating with law enforcement?", "Yes - Documented/Informal/No"),
        
        # Additional cross-integration questions
        ("Q129", "Asset Management", "Is incident management integrated with asset management (A.5.9)?", "Yes/Partial/No"),
        ("Q130", "Access Control", "Is incident management integrated with access control processes (A.5.15-18)?", "Yes/Partial/No"),
        ("Q131", "Change Management", "Is incident management integrated with change management (A.8.32)?", "Yes/Partial/No"),
        ("Q132", "Vulnerability Management", "Is incident management integrated with vulnerability management (A.8.8)?", "Yes/Partial/No"),
        ("Q133", "Patch Management", "Is incident management integrated with patch management (A.8.19)?", "Yes/Partial/No"),
        ("Q134", "Risk Assessment", "Are incident trends fed back into risk assessment processes (A.5.8)?", "Yes/Partial/No"),
        ("Q135", "Audit & Compliance", "Are incident records available for audit and compliance reviews?", "Yes/Partial/No"),
    ]

    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(D{row}="","",IF(OR(D{row}="No", D{row}="Informal"), "Yes", "No"))'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12


    # Answer column DV (D column) — provides dropdown for gap-relevant answers
    from openpyxl.worksheet.datavalidation import DataValidation as _DV
    _dv_answer = _DV(type="list", formula1='"Yes,No,Partial,Informal,In-Progress,N/A"', allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(_dv_answer)
    for _r in range(5, row):
        _dv_answer.add(f"D{_r}")

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 8: GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create the Gap Analysis sheet."""
    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "GAP ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A2"] = f"{CONTROL_REF} | Gap Prioritisation and Remediation Planning (50 Gap Capacity)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:J2")

    # Column headers
    row = 4
    headers = ["Gap ID", "Assessment Section", "Gap Description", "Risk Level", "Current State", "Target State", "Remediation Action", "Owner", "Target Date", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Sample row (F2F2F2 grey) — 1 example gap to guide assessors
    row = 5
    sample_gap = ("GAP-001", "Governance", "No RACI matrix documented — sample entry, replace with actual gaps",
                  "High", "Roles informally understood", "RACI matrix created and approved",
                  "Document RACI matrix, review with CSIRT, obtain CISO sign-off", "CSIRT Manager", "", "Open")
    _f2f2f2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col_idx, value in enumerate(sample_gap, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        cell.fill = _f2f2f2
        cell.border = styles["border"]
        cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
    row += 1

    # 50 empty FFFFCC rows for gap entries — NO pre-filled IDs
    for i in range(50):
        for col_idx in range(1, 11):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1

    # Summary section
    row += 2
    ws[f"A{row}"] = "GAP SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f"A{row}:D{row}")

    row += 1
    summary_labels = ["Total Gaps:", "Critical:", "High:", "Medium:", "Low:", "", "Open:", "In Progress:", "Resolved:", "Accepted:"]
    for label in summary_labels:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        if label:  # Don't create formula for empty row
            if label == "Total Gaps:":
                ws[f"B{row}"] = "=COUNTA(A6:A55)"
            elif label in ["Critical:", "High:", "Medium:", "Low:"]:
                ws[f"B{row}"] = f'=COUNTIF(D6:D55, "{label[:-1]}")'
            elif label in ["Open:", "In Progress:", "Resolved:", "Accepted:"]:
                ws[f"B{row}"] = f'=COUNTIF(J6:J55, "{label[:-1]}")'
            ws[f"B{row}"].fill = styles["calculated_cell"]["fill"]
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 9: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register(ws):
    """Create the standard Evidence Register sheet (gold standard)."""
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _input = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = _border

    # Row 3: empty separator
    for c in range(1, 9):
        ws.cell(row=3, column=c).border = _border

    columns = [
        ("Evidence ID", 14), ("Evidence Type", 20), ("Description", 45),
        ("Related Control / Section", 28), ("Collection Date (DD.MM.YYYY)", 22),
        ("Storage Location / Reference", 38), ("Collected By", 22), ("Status", 14),
    ]
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # F2F2F2 sample row
    sample_data = ["EV-001", "Document", "Sample evidence entry — replace with actual evidence",
                   "All Controls", "01.01.2026", "SharePoint/ISMS/Evidence/", "ISMS Team", "Active"]
    for col_idx, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
        cell.fill = _grey_sample
        cell.border = _border

    dv_status = DataValidation(
        type="list",
        formula1='"Active,Archived,Superseded,Pending Review"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    # 100 FFFFCC empty rows
    for r in range(6, 106):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = _input
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 10: DASHBOARD SHEET
# ============================================================================

def create_dashboard(ws, styles):
    """Summary Dashboard — Gold Standard TABLE 1/2/3."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=1, column=_c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: CONTROL_REF subtitle
    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=2, column=_c).border = border

    # TABLE 1 banner (row 3)
    ws.merge_cells("A3:G3")
    ws["A3"] = "COMPLIANCE ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=3, column=_c).border = border

    # TABLE 1 column headers (row 4)
    _t1_headers = ['Assessment Area', 'Questions Answered', 'No Gap', 'Gap Identified', 'N/A', 'Target', 'Compliance %']
    for _c, _h in enumerate(_t1_headers, 1):
        _cell = ws.cell(row=4, column=_c, value=_h)
        _cell.font = Font(bold=True, size=10, color="000000", name="Calibri")
        _cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _cell.border = border
    ws.row_dimensions[4].height = 30

    # TABLE 1 area rows (rows 5+)
    ws.cell(row=5, column=1).value = 'Governance'
    ws.cell(row=5, column=2).value = "=COUNTA('Governance Assessment'!D5:D100)"
    ws.cell(row=5, column=3).value = "=COUNTIF('Governance Assessment'!G5:G100,\"No\")"
    ws.cell(row=5, column=4).value = "=COUNTIF('Governance Assessment'!G5:G100,\"Yes\")"
    ws.cell(row=5, column=5).value = "=COUNTIF('Governance Assessment'!D5:D100,\"N/A\")"
    ws.cell(row=5, column=6).value = '90%'
    ws.cell(row=5, column=7).value = "=IFERROR(C5/(C5+D5)*100,0)"
    ws.cell(row=5, column=7).number_format = "0.0"
    for _c in range(1, 8):
        _cell = ws.cell(row=5, column=_c)
        _cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        _cell.border = border
        _cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=6, column=1).value = 'Organisational Structure'
    ws.cell(row=6, column=2).value = "=COUNTA('Organisational Structure'!D5:D100)"
    ws.cell(row=6, column=3).value = "=COUNTIF('Organisational Structure'!G5:G100,\"No\")"
    ws.cell(row=6, column=4).value = "=COUNTIF('Organisational Structure'!G5:G100,\"Yes\")"
    ws.cell(row=6, column=5).value = "=COUNTIF('Organisational Structure'!D5:D100,\"N/A\")"
    ws.cell(row=6, column=6).value = '85%'
    ws.cell(row=6, column=7).value = "=IFERROR(C6/(C6+D6)*100,0)"
    ws.cell(row=6, column=7).number_format = "0.0"
    for _c in range(1, 8):
        _cell = ws.cell(row=6, column=_c)
        _cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        _cell.border = border
        _cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=7, column=1).value = 'Training & Competency'
    ws.cell(row=7, column=2).value = "=COUNTA('Training & Competency'!D5:D100)"
    ws.cell(row=7, column=3).value = "=COUNTIF('Training & Competency'!G5:G100,\"No\")"
    ws.cell(row=7, column=4).value = "=COUNTIF('Training & Competency'!G5:G100,\"Yes\")"
    ws.cell(row=7, column=5).value = "=COUNTIF('Training & Competency'!D5:D100,\"N/A\")"
    ws.cell(row=7, column=6).value = '90%'
    ws.cell(row=7, column=7).value = "=IFERROR(C7/(C7+D7)*100,0)"
    ws.cell(row=7, column=7).number_format = "0.0"
    for _c in range(1, 8):
        _cell = ws.cell(row=7, column=_c)
        _cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        _cell.border = border
        _cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=8, column=1).value = 'Tools & Technology'
    ws.cell(row=8, column=2).value = "=COUNTA('Tools & Technology'!D5:D100)"
    ws.cell(row=8, column=3).value = "=COUNTIF('Tools & Technology'!G5:G100,\"No\")"
    ws.cell(row=8, column=4).value = "=COUNTIF('Tools & Technology'!G5:G100,\"Yes\")"
    ws.cell(row=8, column=5).value = "=COUNTIF('Tools & Technology'!D5:D100,\"N/A\")"
    ws.cell(row=8, column=6).value = '80%'
    ws.cell(row=8, column=7).value = "=IFERROR(C8/(C8+D8)*100,0)"
    ws.cell(row=8, column=7).number_format = "0.0"
    for _c in range(1, 8):
        _cell = ws.cell(row=8, column=_c)
        _cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        _cell.border = border
        _cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=9, column=1).value = 'Integration'
    ws.cell(row=9, column=2).value = "=COUNTA('Integration Assessment'!D5:D100)"
    ws.cell(row=9, column=3).value = "=COUNTIF('Integration Assessment'!G5:G100,\"No\")"
    ws.cell(row=9, column=4).value = "=COUNTIF('Integration Assessment'!G5:G100,\"Yes\")"
    ws.cell(row=9, column=5).value = "=COUNTIF('Integration Assessment'!D5:D100,\"N/A\")"
    ws.cell(row=9, column=6).value = '85%'
    ws.cell(row=9, column=7).value = "=IFERROR(C9/(C9+D9)*100,0)"
    ws.cell(row=9, column=7).number_format = "0.0"
    for _c in range(1, 8):
        _cell = ws.cell(row=9, column=_c)
        _cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        _cell.border = border
        _cell.alignment = Alignment(horizontal="left", vertical="center")

    # TABLE 1 TOTAL row (row 10)
    ws.cell(row=10, column=1).value = "TOTAL"
    ws.cell(row=10, column=2).value = "=SUM(B5:B9)"
    ws.cell(row=10, column=3).value = "=SUM(C5:C9)"
    ws.cell(row=10, column=4).value = "=SUM(D5:D9)"
    ws.cell(row=10, column=5).value = "=SUM(E5:E9)"
    ws.cell(row=10, column=6).value = "—"
    ws.cell(row=10, column=7).value = "=IFERROR(AVERAGE(G5:G9),0)"
    ws.cell(row=10, column=7).number_format = "0.0"
    for _c in range(1, 8):
        _cell = ws.cell(row=10, column=_c)
        _cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _cell.font = Font(bold=True, name="Calibri")
        _cell.border = border
        _cell.alignment = Alignment(horizontal="left", vertical="center")
    _total_row = 10

    # TABLE 2
    _t2_row = _total_row + 2
    ws.merge_cells(f"A{_t2_row}:G{_t2_row}")
    ws[f"A{_t2_row}"] = 'FRAMEWORK HEALTH METRICS'
    ws[f"A{_t2_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{_t2_row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{_t2_row}"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=_t2_row, column=_c).border = border
    _t2_row += 1
    # TABLE 2 col headers
    _t2_hdrs = ["Metric", "Value", "Target"]
    for _c, _h in enumerate(_t2_hdrs, 1):
        _cell = ws.cell(row=_t2_row, column=_c, value=_h)
        _cell.font = Font(bold=True, size=10, color="000000", name="Calibri")
        _cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _cell.alignment = Alignment(horizontal="center", vertical="center")
        _cell.border = border
    _t2_row += 1
    # TABLE 2 metric rows
    _t2_metrics = [('CSIRT Staffing (FTE)', "='Organisational Structure'!D29", '≥3 FTE'), ('Coverage Model', "='Organisational Structure'!D28", '24x7'), ('Training Completion Rate', "='Training & Competency'!D58", '≥90%'), ('Tabletop Exercise Frequency', "='Training & Competency'!D61", 'Quarterly'), ('Critical Gaps Identified', "='Gap Analysis'!B72", '0'), ('High Priority Gaps', "='Gap Analysis'!B73", '<5')]
    for _label, _formula, _target in _t2_metrics:
        ws.cell(row=_t2_row, column=1).value = _label
        ws.cell(row=_t2_row, column=2).value = _formula
        ws.cell(row=_t2_row, column=3).value = _target
        for _c in range(1, 4):
            _cell = ws.cell(row=_t2_row, column=_c)
            _cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            _cell.border = border
            _cell.alignment = Alignment(horizontal="left", vertical="center")
        _t2_row += 1
    _t3_start = _t2_row

    # TABLE 3
    _t3_row = _t3_start + 1
    ws.merge_cells(f"A{_t3_row}:G{_t3_row}")
    ws[f"A{_t3_row}"] = 'CRITICAL FINDINGS'
    ws[f"A{_t3_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{_t3_row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{_t3_row}"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=_t3_row, column=_c).border = border
    _t3_row += 1
    # TABLE 3 col headers
    _t3_hdrs = ["Critical Finding", "Status", "Severity"]
    for _c, _h in enumerate(_t3_hdrs, 1):
        _cell = ws.cell(row=_t3_row, column=_c, value=_h)
        _cell.font = Font(bold=True, size=10, color="000000", name="Calibri")
        _cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _cell.alignment = Alignment(horizontal="center", vertical="center")
        _cell.border = border
    _t3_row += 1
    _t3_findings = [('Governance compliance <70%', '=IF(G5<70,"[!] Below Target","[OK]")', 'Critical'), ('Organisational Structure <70%', '=IF(G6<70,"[!] Below Target","[OK]")', 'Critical'), ('Training & Competency <70%', '=IF(G7<70,"[!] Below Target","[OK]")', 'High'), ('Tools & Technology <70%', '=IF(G8<70,"[!] Below Target","[OK]")', 'High'), ('Integration <70%', '=IF(G9<70,"[!] Below Target","[OK]")', 'High'), ('Overall framework maturity <70%', '=IF(G10<70,"[!] Below Target","[OK]")', 'Critical'), ('Critical gaps exist', '=IF(ISNUMBER(\'Gap Analysis\'!B72),IF(\'Gap Analysis\'!B72>0,"[!] Gaps Present","[OK]"),"")', 'Critical')]
    _t3_sev_fills = {'Critical': 'FFC7CE', 'High': 'FFEB9C', 'Low': 'C6EFCE'}
    for _label, _formula, _severity in _t3_findings:
        _fill_color = _t3_sev_fills.get(_severity, "FFFFCC")
        ws.cell(row=_t3_row, column=1).value = _label
        ws.cell(row=_t3_row, column=2).value = _formula
        ws.cell(row=_t3_row, column=3).value = _severity
        for _c in range(1, 4):
            _cell = ws.cell(row=_t3_row, column=_c)
            _cell.fill = PatternFill(start_color=_fill_color, end_color=_fill_color, fill_type="solid")
            _cell.border = border
            _cell.alignment = Alignment(horizontal="left", vertical="center")
        _t3_row += 1

    # Column widths + freeze
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A4"


def create_approval_sheet(ws):
    """Create the Approval Sign-Off sheet — Gold Standard."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Summary fields (rows 4-8)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", ""),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    # GS-AS-015: Overall Compliance Rate — must reference Summary Dashboard
    ws["B6"].value = "=IFERROR(AVERAGE('Summary Dashboard'!G5:G9),\"\")"
    ws["B6"].number_format = "0.0%"

    # Assessment Status dropdown (row 7)
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # 3 Approver sections (start at row 11)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: MAIN EXECUTION
# ============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        ws.data_validations.dataValidation = [
            dv for dv in list(ws.data_validations.dataValidation)
            if dv.sqref
        ]


def main():
    """Main execution function - orchestrates workbook creation."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.5.24-28.S1 - Framework Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.5.24: Planning & Preparation")
    logger.info("=" * 80)

    wb = create_workbook()
    styles = _STYLES

    logger.info("\n[1/10] Creating Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"])

    logger.info("[2/10] Creating Governance Assessment sheet (25 questions)...")
    create_governance_assessment(wb["Governance Assessment"], styles)

    logger.info("[3/10] Creating Organisational Structure sheet (30 questions)...")
    create_organisational_structure(wb["Organisational Structure"], styles)

    logger.info("[4/10] Creating Training & Competency sheet (25 questions)...")
    create_training_competency(wb["Training & Competency"], styles)

    logger.info("[5/10] Creating Tools & Technology sheet (30 questions)...")
    create_tools_technology(wb["Tools & Technology"], styles)

    logger.info("[6/10] Creating Integration Assessment sheet (25 questions)...")
    create_integration_assessment(wb["Integration Assessment"], styles)

    logger.info("[7/10] Creating Gap Analysis sheet...")
    create_gap_analysis(wb["Gap Analysis"], styles)

    logger.info("[8/10] Creating Evidence Register sheet...")
    create_evidence_register(wb["Evidence Register"])

    logger.info("[9/10] Creating Dashboard sheet...")
    create_dashboard(wb["Summary Dashboard"], styles)

    logger.info("[10/10] Creating Approval Sign-Off sheet...")
    create_approval_sheet(wb["Approval Sign-Off"])

    finalize_validations(wb)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    output_path = _wkbk_dir / OUTPUT_FILENAME
    wb.save(output_path)
    logger.info(f"\nSUCCESS: {_wkbk_dir / OUTPUT_FILENAME}")
    logger.info("\nWorkbook Structure:")
    logger.info("  • 10 sheets (Instructions through Approval)")
    logger.info("  • 135 assessment questions (25+30+25+30+25)")
    logger.info("  • 60 gap analysis capacity")
    logger.info("  • 100 evidence register capacity")
    logger.info("  • Automated maturity scoring (Levels 1-5)")
    logger.info("\nNext steps:")
    logger.info("  1) Complete document information in Instructions & Legend")
    logger.info("  2) Conduct stakeholder interviews (CSIRT, Legal, HR, IT Ops)")
    logger.info("  3) Fill yellow cells in assessment sheets (Governance through Integration)")
    logger.info("  4) Review automated gap identification in Gap Analysis")
    logger.info("  5) Collect and link evidence in Evidence Register")
    logger.info("  6) Review Dashboard for maturity scoring")
    logger.info("  7) Complete Approval Sign-Off workflow")
    logger.info("\nEstimated completion time: 8-12 hours")
    logger.info("\n" + "=" * 80)


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
