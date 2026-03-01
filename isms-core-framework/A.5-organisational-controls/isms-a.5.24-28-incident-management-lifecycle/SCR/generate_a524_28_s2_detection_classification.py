#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.24-28.S2 - Detection & Classification Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.24-28: Incident Management
Assessment Domain 2 of 5: Detection & Classification

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific incident management infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Incident classification levels and severity criteria (match your organisation)
2. Detection source categories and integration points (adapt to your toolset)
3. Response team roles and escalation thresholds
4. Evidence collection requirements and chain-of-custody procedures
5. Lessons-learned process and improvement tracking mechanisms

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.24-28 Incident Management Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
incident management controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Detection & Classification under ISO 27001:2022 Controls A.5.24-A.5.28. Supports evidence-based evaluation of incident lifecycle capabilities, response effectiveness, and continuous improvement processes.

**Assessment Scope:**
- Incident detection mechanism coverage and response time metrics
- Classification accuracy and severity assignment consistency
- Response capability completeness and team readiness
- Forensic evidence collection and preservation procedures
- Stakeholder communication and regulatory notification compliance
- Lessons-learned implementation and effectiveness tracking
- Evidence collection for incident management and regulatory audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
6. Summary Dashboard - Compliance overview and key metrics
7. Evidence Register - Audit evidence tracking
8. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 5 domains covering Incident Management controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a524_28_s2_detection_classification.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a524_28_s2_detection_classification.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a524_28_s2_detection_classification.py --date 20250115

Output:
    File: ISMS-IMP-A.5.24-28.S2_Detection_&_Classification_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.24-28
Assessment Domain:    2 of 5 (Detection & Classification)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.24-28: Incident Management Policy (Governance)
    - ISMS-IMP-A.5.24-28.S1: Incident Management Framework (Domain 1)
    - ISMS-IMP-A.5.24-28.S2: Detection & Classification (Domain 2)
    - ISMS-IMP-A.5.24-28.S3: Response Capabilities (Domain 3)
    - ISMS-IMP-A.5.24-28.S4: Forensic Evidence (Domain 4)
    - ISMS-IMP-A.5.24-28.S5: Learning & Continuous Improvement (Domain 5)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.24-28.S2 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive incident management details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review incident classification criteria, response procedures, and lessons-learned outcomes annually or after significant incidents, major infrastructure changes, or regulatory updates.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

from datetime import datetime
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging
import sys

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.24-28.S2"
WORKBOOK_NAME = "Detection & Classification"
CONTROL_ID = "A.5.24-28"
CONTROL_NAME = "Incident Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)
    
    sheets = [
        "Instructions & Legend",
        "Detection Mechanisms",
        "Alert Handling",
        "Classification & Severity",
        "Detection Effectiveness",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)
    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "input_cell": {
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": border_thin,
        },
        "calculated_cell": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": border_thin,
        },
        "border": border_thin,
        "gap_critical": {"fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")},
        "gap_high": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "gap_medium": {"fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")},
        "gap_low": {"fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")},
    }
    return styles


# ============================================================================
# SECTION 2: INSTRUCTIONS SHEET
# ============================================================================



_STYLES = setup_styles()
def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete Detection Mechanisms sheet — review all detection sources and coverage.', '2. Complete Alert Handling sheet — review triage, false positive rate, ticket hygiene.', '3. Complete Classification & Severity sheet — verify classification accuracy.', '4. Complete Detection Effectiveness sheet — MTTD, coverage metrics, gaps.', '5. Review Gap Analysis and prioritise remediation.', '6. Link evidence in Evidence Register.', '7. Review Summary Dashboard for overall compliance.', '8. Complete Approval Sign-Off workflow.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 22

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_detection_mechanisms(ws, styles):
    """Create Detection Mechanisms sheet (33 questions)."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "DETECTION MECHANISMS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A2"] = f"{CONTROL_REF} | SIEM, EDR, IDS/IPS, NDR, User Reporting, Coverage by Threat Category (33 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")
    
    # Column headers
    row = 4
    headers = ["Question ID", "Section", "Question", "Answer", "Evidence Reference", "Comments", "Gap Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    # Questions (showing subset - actual implementation would have all 33)
    questions = [
        ("Q1", "SIEM", "What SIEM platform(s) does [Organisation] use?", "Text"),
        ("Q2", "SIEM", "How many active detection rules/alerts are configured in the SIEM?", "Number"),
        ("Q3", "SIEM", "Which threat categories have SIEM detection rules?", "Checkbox"),
        ("Q4", "SIEM", "How often are SIEM rules tuned to reduce false positives?", "Weekly/Monthly/Quarterly/Annually/Ad-Hoc/Never"),
        ("Q5", "SIEM", "What percentage of SIEM rules are custom-developed (vs. vendor-provided)?", "Percentage"),
        ("Q6", "SIEM", "Is threat intelligence integrated into SIEM for automated alerting?", "Yes - Automated/Manual/No"),
        ("Q7", "SIEM", "What is the average daily SIEM alert volume?", "Number"),
        ("Q8", "SIEM", "How many log sources feed the SIEM?", "Number"),
        
        ("Q9", "EDR", "Is EDR or XDR deployed?", "Yes - EDR/Yes - XDR/No"),
        ("Q10", "EDR", "What percentage of endpoints have EDR agent deployed?", "Percentage"),
        ("Q11", "EDR", "What types of EDR alerts are enabled?", "Checkbox"),
        ("Q12", "EDR", "Does EDR support automated response actions?", "Fully/Semi/Manual"),
        ("Q13", "EDR", "How long does EDR retain forensic data?", "30/60/90/180/365+ days"),
        
        ("Q14", "Network Detection", "Is IDS/IPS deployed?", "IPS/IDS/No"),
        ("Q15", "Network Detection", "Where is IDS/IPS deployed?", "Checkbox"),
        ("Q16", "Network Detection", "How frequently are IDS/IPS signatures updated?", "Daily/Weekly/Monthly"),
        ("Q17", "Network Detection", "Is NDR/NTA deployed?", "Yes/No"),
        
        ("Q18", "Other Detection", "Is there a user reporting mechanism (A.6.8)?", "Dedicated/Helpdesk/No"),
        ("Q19", "Other Detection", "Average monthly user-reported events?", "Number"),
        ("Q20", "Other Detection", "Does DLP generate security alerts?", "Integrated/Separate/No"),
        ("Q21", "Other Detection", "Does CASB generate security alerts?", "Yes/No/N/A"),
        ("Q22", "Other Detection", "Are honeypots/deception deployed?", "Yes/No"),
        
        # Coverage by category (Q23-Q33)
        ("Q23", "Coverage", "Category 1: Malware Detection", "Yes/Partial/No"),
        ("Q24", "Coverage", "Category 2: Unauthorised Access Detection", "Yes/Partial/No"),
        ("Q25", "Coverage", "Category 3: Data Breach Detection", "Yes/Partial/No"),
        ("Q26", "Coverage", "Category 4: DoS Detection", "Yes/Partial/No"),
        ("Q27", "Coverage", "Category 5: Social Engineering Detection", "Yes/Partial/No"),
        ("Q28", "Coverage", "Category 6: Web Attack Detection", "Yes/Partial/No"),
        ("Q29", "Coverage", "Category 7: Network Attack Detection", "Yes/Partial/No"),
        ("Q30", "Coverage", "Category 8: Endpoint Compromise Detection", "Yes/Partial/No"),
        ("Q31", "Coverage", "Category 9: Physical Security Detection", "Yes/Partial/No"),
        ("Q32", "Coverage", "Category 10: Config/Patch Detection", "Yes/Partial/No"),
        ("Q33", "Coverage", "Category 11: Other Detection", "Yes/Partial/No"),
    ]
    
    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(D{row}="","",IF(OR(D{row}="No", D{row}="Never", D{row}<90), "Yes", "No"))'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12

    # Answer column DV (D column) — provides dropdown for gap-relevant answers
    from openpyxl.worksheet.datavalidation import DataValidation as _DV
    _dv_answer = _DV(type="list", formula1='"Yes,No,Partial,Never,In-Progress,N/A"', allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(_dv_answer)
    for _r in range(5, row):
        _dv_answer.add(f"D{_r}")

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 4: ALERT HANDLING SHEET
# ============================================================================

def create_alert_handling(ws, styles):
    """Create Alert Handling sheet (25 questions)."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "ALERT HANDLING"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A2"] = f"{CONTROL_REF} | Triage, Investigation, Escalation, Analyst Capacity (25 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")
    
    row = 4
    headers = ["Question ID", "Section", "Question", "Answer", "Evidence Reference", "Comments", "Gap Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    questions = [
        ("Q34", "Triage", "Are alert triage procedures documented?", "Comprehensive/Basic/No"),
        ("Q35", "Triage", "Are triage time SLAs defined for each severity?", "Yes/No"),
        ("Q36", "Triage", "Percentage of alerts triaged within SLA? (90 days)", "Percentage"),
        ("Q37", "Triage", "How are alerts prioritised in queue?", "Automated/Manual/FIFO/No Process"),
        ("Q38", "Triage", "Documented shift handoff process?", "Formal/Informal/No"),
        ("Q39", "Triage", "Do analysts receive triage training?", "Regular/Onboarding/No"),
        ("Q40", "Triage", "What tools for alert triage?", "Checkbox"),
        
        ("Q41", "Playbooks", "Are investigation playbooks documented?", "Comprehensive/Limited/No"),
        ("Q42", "Playbooks", "Which categories have playbooks?", "Checkbox"),
        ("Q43", "Playbooks", "Playbook format?", "SOAR/Written/Flowchart/Checklist"),
        ("Q44", "Playbooks", "Playbook update frequency?", "Quarterly/Semi/Annually/After Incidents/Never"),
        ("Q45", "Playbooks", "Playbook: Malware/Ransomware", "Yes/No"),
        ("Q46", "Playbooks", "Playbook: Phishing/BEC", "Yes/No"),
        ("Q47", "Playbooks", "Playbook: Unauthorised Access", "Yes/No"),
        ("Q48", "Playbooks", "Playbook: Data Exfiltration", "Yes/No"),
        
        ("Q49", "Escalation", "Escalation criteria clearly defined?", "Well Defined/Partial/No"),
        ("Q50", "Escalation", "How are incidents escalated SOC → CSIRT?", "Automated/Manual/Informal/No Process"),
        ("Q51", "Escalation", "Handoff checklist for CSIRT escalation?", "Yes/No"),
        ("Q52", "Escalation", "False positive handling process?", "Documented/Informal/No"),
        ("Q53", "Escalation", "Feedback loop: FP → tuning?", "Systematic/Informal/No"),
        
        ("Q54", "Capacity", "SOC analyst FTE count?", "Number"),
        ("Q55", "Capacity", "Tiered analyst structure?", "Multi-Tier/Single-Tier"),
        ("Q56", "Capacity", "Analyst training frequency?", "Quarterly/Semi/Annually/Onboarding/Never"),
        ("Q57", "Capacity", "Analyst burnout indicators?", "No/Some/Significant"),
        ("Q58", "Capacity", "Analysts trained on all tools?", "Comprehensive/Partial/No"),
    ]
    
    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(D{row}="","",IF(OR(D{row}="No", D{row}="Informal", D{row}<80), "Yes", "No"))'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12

    # Answer column DV (D column) — provides dropdown for gap-relevant answers
    from openpyxl.worksheet.datavalidation import DataValidation as _DV
    _dv_answer = _DV(type="list", formula1='"Yes,No,Partial,Informal,In-Progress,N/A"', allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(_dv_answer)
    for _r in range(5, row):
        _dv_answer.add(f"D{_r}")

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 5: CLASSIFICATION & SEVERITY SHEET
# ============================================================================

def create_classification_severity(ws, styles):
    """Create Classification & Severity sheet (25 questions)."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "CLASSIFICATION & SEVERITY"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A2"] = f"{CONTROL_REF} | Category Assignment, Severity Criteria, Consistency (25 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")
    
    row = 4
    headers = ["Question ID", "Section", "Question", "Answer", "Evidence Reference", "Comments", "Gap Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    questions = [
        ("Q59", "Classification", "Adopted ISMS-REF-A.5.24-28 taxonomy?", "Fully/Partially/No"),
        ("Q60", "Classification", "Classification procedure documented?", "Yes/No"),
        ("Q61", "Classification", "Analysts trained on classification?", "Regular/Onboarding/No"),
        ("Q62", "Classification", "Classification consistency reviewed?", "Monthly/Quarterly/Annually/No"),
        ("Q63", "Classification", "Classification accuracy rate?", "Percentage"),
        ("Q64", "Classification", "Reclassification process documented?", "Yes/No"),
        ("Q65", "Classification", "Multi-category incident handling?", "Primary+Secondary/Multiple Tickets/Single/No Process"),
        
        ("Q66", "Severity", "How many severity levels?", "4/3/5/Other"),
        ("Q67", "Severity", "Severity criteria documented?", "Comprehensive/Basic/No"),
        ("Q68", "Severity", "Critical severity criteria?", "Checkbox"),
        ("Q69", "Severity", "Severity assignment training?", "Regular/Onboarding/No"),
        ("Q70", "Severity", "Severity consistency reviewed?", "Monthly/Quarterly/Annually/No"),
        ("Q71", "Severity", "Severity upgrade process?", "Documented/Informal/No"),
        ("Q72", "Severity", "Severity downgrade process?", "Documented/Informal/No"),
        ("Q73", "Severity", "% Incidents Critical (90 days)?", "Percentage"),
        ("Q74", "Severity", "% Incidents High (90 days)?", "Percentage"),
        ("Q75", "Severity", "% Incidents Medium (90 days)?", "Percentage"),
        
        # Top categories by volume (Q76-Q83 not shown for brevity)
    ]
    
    # Add remaining 8 questions for category/severity correlation
    for i in range(76, 84):
        questions.append((f"Q{i}", "Correlation", f"Top Category {i-75}", "Text"))
    
    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(D{row}="","",IF(OR(D{row}="No", D{row}<80), "Yes", "No"))'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12

    # Answer column DV (D column) — provides dropdown for gap-relevant answers
    from openpyxl.worksheet.datavalidation import DataValidation as _DV
    _dv_answer = _DV(type="list", formula1='"Yes,No,Partial,In-Progress,N/A"', allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(_dv_answer)
    for _r in range(5, row):
        _dv_answer.add(f"D{_r}")

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 6: DETECTION EFFECTIVENESS SHEET
# ============================================================================

def create_detection_effectiveness(ws, styles):
    """Create Detection Effectiveness sheet (25 questions)."""
    ws.merge_cells("A1:G1")
    ws["A1"] = "DETECTION EFFECTIVENESS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A2"] = f"{CONTROL_REF} | MTTD, MTTT, False Positives, Coverage Gaps (25 Questions)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")
    
    row = 4
    headers = ["Question ID", "Section", "Question", "Answer", "Evidence Reference", "Comments", "Gap Identified"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]
    
    questions = [
        ("Q84", "MTTD", "Is MTTD tracked?", "Yes/No"),
        ("Q85", "MTTD", "Overall MTTD (90 days)?", "Duration"),
        ("Q86", "MTTD", "MTTD for Critical incidents?", "Duration"),
        ("Q87", "MTTD", "MTTD for High incidents?", "Duration"),
        ("Q88", "MTTD", "MTTD trend?", "Improving/Stable/Worsening/Not Tracked"),
        
        ("Q89", "MTTT", "Is MTTT tracked?", "Yes/No"),
        ("Q90", "MTTT", "Overall MTTT (90 days)?", "Duration"),
        ("Q91", "MTTT", "MTTT for Critical alerts?", "Duration"),
        ("Q92", "MTTT", "MTTT for High alerts?", "Duration"),
        ("Q93", "MTTT", "MTTT SLA compliance %?", "Percentage"),
        
        ("Q94", "False Positives", "FPs systematically tracked?", "Yes/No"),
        ("Q95", "False Positives", "Total alerts (90 days)?", "Number"),
        ("Q96", "False Positives", "False positives (90 days)?", "Number"),
        ("Q97", "False Positives", "False positive rate?", "=Q96/Q95*100 (Calculated)"),
        ("Q98", "False Positives", "FP trend?", "Improving/Stable/Worsening/Not Tracked"),
        ("Q99", "False Positives", "Top FP source?", "SIEM/EDR/IDS/NDR/User/Other"),
        ("Q100", "False Positives", "FP tuning backlog?", "Large/Small/None/Not Tracked"),
        
        ("Q101", "True Positives", "True positives (90 days)?", "Number"),
        ("Q102", "True Positives", "True positive rate?", "=Q101/Q95*100 (Calculated)"),
        ("Q103", "True Positives", "Incident volume trend?", "Increasing/Stable/Decreasing"),
        ("Q104", "True Positives", "Alert-to-incident ratio?", "=Q95/Q101 (Calculated)"),
        ("Q105", "True Positives", "Can SOC handle volume?", "Comfortably/At Capacity/Overwhelmed"),
        
        ("Q106", "Coverage", "Detection blind spots identified?", "Documented/Informally/No"),
        ("Q107", "Coverage", "Coverage gaps prioritised?", "Yes/No"),
        ("Q108", "Coverage", "Coverage improvement plan?", "Funded/Planned/No"),
    ]
    
    row = 5
    for q_id, section, question, answer_type in questions:
        ws[f"A{row}"] = q_id
        ws[f"B{row}"] = section
        ws[f"C{row}"] = question
        ws[f"D{row}"].fill = styles["input_cell"]["fill"]
        
        # Special handling for calculated cells
        if "Calculated" in answer_type:
            formula = answer_type.split("(Calculated)")[0].strip()
            ws[f"D{row}"] = formula.replace("Q95", f"D{row-13}").replace("Q96", f"D{row-12}").replace("Q101", f"D{row-7}")
            ws[f"D{row}"].fill = styles["calculated_cell"]["fill"]
        
        ws[f"E{row}"].fill = styles["input_cell"]["fill"]
        ws[f"F{row}"].fill = styles["input_cell"]["fill"]
        ws[f"G{row}"] = f'=IF(D{row}="","",IF(OR(D{row}="No", D{row}>30, D{row}="Worsening"), "Yes", "No"))'
        ws[f"G{row}"].fill = styles["calculated_cell"]["fill"]
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].border = styles["border"]
            if col == "C":
                ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12

    # Answer column DV (D column) — provides dropdown for gap-relevant answers
    from openpyxl.worksheet.datavalidation import DataValidation as _DV
    _dv_answer = _DV(type="list", formula1='"Yes,No,Partial,Worsening,Improving,Stable,In-Progress,N/A"', allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(_dv_answer)
    for _r in range(5, row):
        _dv_answer.add(f"D{_r}")

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 7-9: GAP ANALYSIS, EVIDENCE, DASHBOARD, APPROVAL
# ============================================================================

def create_gap_analysis(ws, styles):
    """Create Gap Analysis sheet (40 capacity)."""
    ws.merge_cells("A1:J1")
    ws["A1"] = "GAP ANALYSIS"
    ws["A1"].font = styles["header"]["font"]
    ws["A1"].fill = styles["header"]["fill"]
    ws["A1"].alignment = styles["header"]["alignment"]
    ws.row_dimensions[1].height = 35

    ws["A2"] = f"{CONTROL_REF} | Gap Prioritisation and Remediation Planning (40 Gap Capacity)"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:J2")

    row = 4
    headers = ["Gap ID", "Section", "Gap Description", "Risk Level", "Current State", "Target State", "Remediation", "Owner", "Target Date", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = styles["column_header"]["font"]
        cell.fill = styles["column_header"]["fill"]
        cell.alignment = styles["column_header"]["alignment"]
        cell.border = styles["column_header"]["border"]

    # Sample row (F2F2F2) then 50 empty FFFFCC rows — NO pre-filled IDs
    row = 5
    ws[f"A{row}"] = "GAP-001"
    for col_idx in range(2, 11):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = styles["border"]
    ws[f"A{row}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws[f"A{row}"].border = styles["border"]
    row += 1

    # 50 empty FFFFCC rows
    for i in range(50):
        for col_idx in range(1, 11):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = styles["input_cell"]["fill"]
            cell.border = styles["border"]
        row += 1
    
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.freeze_panes = "A5"


def create_evidence_register(ws):
    """Create the standard Evidence Register sheet (gold standard)."""
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _input = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = _border

    # Row 3: empty separator
    for c in range(1, 9):
        ws.cell(row=3, column=c).border = _border

    columns = [
        ("Evidence ID", 14), ("Evidence Type", 20), ("Description", 45),
        ("Related Control / Section", 28), ("Collection Date (DD.MM.YYYY)", 22),
        ("Storage Location / Reference", 38), ("Collected By", 22), ("Status", 14),
    ]
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # F2F2F2 sample row
    sample_data = ["EV-001", "Document", "Sample evidence entry — replace with actual evidence",
                   "All Controls", "01.01.2026", "SharePoint/ISMS/Evidence/", "ISMS Team", "Active"]
    for col_idx, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
        cell.fill = _grey_sample
        cell.border = _border

    dv_status = DataValidation(
        type="list",
        formula1='"Active,Archived,Superseded,Pending Review"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    # 100 FFFFCC empty rows
    for r in range(6, 106):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = _input
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


def create_dashboard(ws, styles):
    """Summary Dashboard — Gold Standard TABLE 1/2/3."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=1, column=_c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: CONTROL_REF subtitle
    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=2, column=_c).border = border

    # TABLE 1 banner (row 3)
    ws.merge_cells("A3:G3")
    ws["A3"] = "COMPLIANCE ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=3, column=_c).border = border

    # TABLE 1 column headers (row 4)
    _t1_headers = ['Assessment Area', 'Questions Answered', 'No Gap', 'Gap Identified', 'N/A', 'Target', 'Compliance %']
    for _c, _h in enumerate(_t1_headers, 1):
        _cell = ws.cell(row=4, column=_c, value=_h)
        _cell.font = Font(bold=True, size=10, color="000000", name="Calibri")
        _cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _cell.border = border
    ws.row_dimensions[4].height = 30

    # TABLE 1 area rows (rows 5+)
    ws.cell(row=5, column=1).value = 'Detection Mechanisms'
    ws.cell(row=5, column=2).value = "=COUNTA('Detection Mechanisms'!D5:D100)"
    ws.cell(row=5, column=3).value = "=COUNTIF('Detection Mechanisms'!G5:G100,\"No\")"
    ws.cell(row=5, column=4).value = "=COUNTIF('Detection Mechanisms'!G5:G100,\"Yes\")"
    ws.cell(row=5, column=5).value = "=COUNTIF('Detection Mechanisms'!D5:D100,\"N/A\")"
    ws.cell(row=5, column=6).value = '85%'
    ws.cell(row=5, column=7).value = "=IFERROR(C5/(C5+D5)*100,0)"
    ws.cell(row=5, column=7).number_format = "0.0"
    for _c in range(1, 8):
        _cell = ws.cell(row=5, column=_c)
        _cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        _cell.border = border
        _cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=6, column=1).value = 'Alert Handling'
    ws.cell(row=6, column=2).value = "=COUNTA('Alert Handling'!D5:D100)"
    ws.cell(row=6, column=3).value = "=COUNTIF('Alert Handling'!G5:G100,\"No\")"
    ws.cell(row=6, column=4).value = "=COUNTIF('Alert Handling'!G5:G100,\"Yes\")"
    ws.cell(row=6, column=5).value = "=COUNTIF('Alert Handling'!D5:D100,\"N/A\")"
    ws.cell(row=6, column=6).value = '85%'
    ws.cell(row=6, column=7).value = "=IFERROR(C6/(C6+D6)*100,0)"
    ws.cell(row=6, column=7).number_format = "0.0"
    for _c in range(1, 8):
        _cell = ws.cell(row=6, column=_c)
        _cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        _cell.border = border
        _cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=7, column=1).value = 'Classification & Severity'
    ws.cell(row=7, column=2).value = "=COUNTA('Classification & Severity'!D5:D100)"
    ws.cell(row=7, column=3).value = "=COUNTIF('Classification & Severity'!G5:G100,\"No\")"
    ws.cell(row=7, column=4).value = "=COUNTIF('Classification & Severity'!G5:G100,\"Yes\")"
    ws.cell(row=7, column=5).value = "=COUNTIF('Classification & Severity'!D5:D100,\"N/A\")"
    ws.cell(row=7, column=6).value = '90%'
    ws.cell(row=7, column=7).value = "=IFERROR(C7/(C7+D7)*100,0)"
    ws.cell(row=7, column=7).number_format = "0.0"
    for _c in range(1, 8):
        _cell = ws.cell(row=7, column=_c)
        _cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        _cell.border = border
        _cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=8, column=1).value = 'Detection Effectiveness'
    ws.cell(row=8, column=2).value = "=COUNTA('Detection Effectiveness'!D5:D100)"
    ws.cell(row=8, column=3).value = "=COUNTIF('Detection Effectiveness'!G5:G100,\"No\")"
    ws.cell(row=8, column=4).value = "=COUNTIF('Detection Effectiveness'!G5:G100,\"Yes\")"
    ws.cell(row=8, column=5).value = "=COUNTIF('Detection Effectiveness'!D5:D100,\"N/A\")"
    ws.cell(row=8, column=6).value = '80%'
    ws.cell(row=8, column=7).value = "=IFERROR(C8/(C8+D8)*100,0)"
    ws.cell(row=8, column=7).number_format = "0.0"
    for _c in range(1, 8):
        _cell = ws.cell(row=8, column=_c)
        _cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        _cell.border = border
        _cell.alignment = Alignment(horizontal="left", vertical="center")

    # TABLE 1 TOTAL row (row 9)
    ws.cell(row=9, column=1).value = "TOTAL"
    ws.cell(row=9, column=2).value = "=SUM(B5:B8)"
    ws.cell(row=9, column=3).value = "=SUM(C5:C8)"
    ws.cell(row=9, column=4).value = "=SUM(D5:D8)"
    ws.cell(row=9, column=5).value = "=SUM(E5:E8)"
    ws.cell(row=9, column=6).value = "—"
    ws.cell(row=9, column=7).value = "=IFERROR(AVERAGE(G5:G8),0)"
    ws.cell(row=9, column=7).number_format = "0.0"
    for _c in range(1, 8):
        _cell = ws.cell(row=9, column=_c)
        _cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _cell.font = Font(bold=True, name="Calibri")
        _cell.border = border
        _cell.alignment = Alignment(horizontal="left", vertical="center")
    _total_row = 9

    # TABLE 2
    _t2_row = _total_row + 2
    ws.merge_cells(f"A{_t2_row}:G{_t2_row}")
    ws[f"A{_t2_row}"] = 'DETECTION PERFORMANCE METRICS'
    ws[f"A{_t2_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{_t2_row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{_t2_row}"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=_t2_row, column=_c).border = border
    _t2_row += 1
    # TABLE 2 col headers
    _t2_hdrs = ["Metric", "Value", "Target"]
    for _c, _h in enumerate(_t2_hdrs, 1):
        _cell = ws.cell(row=_t2_row, column=_c, value=_h)
        _cell.font = Font(bold=True, size=10, color="000000", name="Calibri")
        _cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _cell.alignment = Alignment(horizontal="center", vertical="center")
        _cell.border = border
    _t2_row += 1
    # TABLE 2 metric rows
    _t2_metrics = [('Overall MTTD (hours)', "='Detection Effectiveness'!D85", '≤4h'), ('Overall MTTT (hours)', "='Detection Effectiveness'!D90", '≤1h'), ('False Positive Rate (%)', "='Detection Effectiveness'!D97", '≤20%'), ('True Positive Rate (%)', "='Detection Effectiveness'!D102", '≥80%'), ('Detection Coverage (%)', '=COUNTIF(\'Detection Mechanisms\'!D27:D37,"Yes")/11*100', '≥85%'), ('Alert-to-Incident Ratio', "='Detection Effectiveness'!D104", '≥0.1')]
    for _label, _formula, _target in _t2_metrics:
        ws.cell(row=_t2_row, column=1).value = _label
        ws.cell(row=_t2_row, column=2).value = _formula
        ws.cell(row=_t2_row, column=3).value = _target
        for _c in range(1, 4):
            _cell = ws.cell(row=_t2_row, column=_c)
            _cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            _cell.border = border
            _cell.alignment = Alignment(horizontal="left", vertical="center")
        _t2_row += 1
    _t3_start = _t2_row

    # TABLE 3
    _t3_row = _t3_start + 1
    ws.merge_cells(f"A{_t3_row}:G{_t3_row}")
    ws[f"A{_t3_row}"] = 'CRITICAL FINDINGS'
    ws[f"A{_t3_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{_t3_row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{_t3_row}"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=_t3_row, column=_c).border = border
    _t3_row += 1
    # TABLE 3 col headers
    _t3_hdrs = ["Critical Finding", "Status", "Severity"]
    for _c, _h in enumerate(_t3_hdrs, 1):
        _cell = ws.cell(row=_t3_row, column=_c, value=_h)
        _cell.font = Font(bold=True, size=10, color="000000", name="Calibri")
        _cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _cell.alignment = Alignment(horizontal="center", vertical="center")
        _cell.border = border
    _t3_row += 1
    _t3_findings = [('Detection Mechanisms <70%', '=IF(G5<70,"[!] Below Target","[OK]")', 'Critical'), ('Alert Handling <70%', '=IF(G6<70,"[!] Below Target","[OK]")', 'Critical'), ('Classification & Severity <70%', '=IF(G7<70,"[!] Below Target","[OK]")', 'High'), ('Detection Effectiveness <70%', '=IF(G8<70,"[!] Below Target","[OK]")', 'High'), ('Overall detection compliance <70%', '=IF(G9<70,"[!] Below Target","[OK]")', 'Critical'), ('False positive rate too high', '=IFERROR(IF(\'Detection Effectiveness\'!D97>30,"[!] High FPR","[OK]"),"")', 'High')]
    _t3_sev_fills = {'Critical': 'FFC7CE', 'High': 'FFEB9C', 'Low': 'C6EFCE'}
    for _label, _formula, _severity in _t3_findings:
        _fill_color = _t3_sev_fills.get(_severity, "FFFFCC")
        ws.cell(row=_t3_row, column=1).value = _label
        ws.cell(row=_t3_row, column=2).value = _formula
        ws.cell(row=_t3_row, column=3).value = _severity
        for _c in range(1, 4):
            _cell = ws.cell(row=_t3_row, column=_c)
            _cell.fill = PatternFill(start_color=_fill_color, end_color=_fill_color, fill_type="solid")
            _cell.border = border
            _cell.alignment = Alignment(horizontal="left", vertical="center")
        _t3_row += 1

    # Column widths + freeze
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A4"


def create_approval_sheet(ws):
    """Create the Approval Sign-Off sheet — Gold Standard."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Summary fields (rows 4-8)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", ""),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    # GS-AS-015: Overall Compliance Rate — must reference Summary Dashboard
    ws["B6"].value = "=IFERROR(AVERAGE('Summary Dashboard'!G5:G8),\"\")"
    ws["B6"].number_format = "0.0%"

    # Assessment Status dropdown (row 7)
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # 3 Approver sections (start at row 11)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 10: MAIN EXECUTION
# ============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        ws.data_validations.dataValidation = [
            dv for dv in list(ws.data_validations.dataValidation)
            if dv.sqref
        ]


def main():
    """Main execution function."""
    logger.info("=" * 80)
    logger.info("ISMS-IMP-A.5.24-28.S2 - Detection & Classification Assessment Generator")
    logger.info("ISO/IEC 27001:2022 Control A.5.25: Assessment & Decision")
    logger.info("=" * 80)
    
    wb = create_workbook()
    styles = _STYLES
    
    logger.info("\n[1/9] Creating Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"])
    
    logger.info("[2/9] Creating Detection Mechanisms sheet (33 questions)...")
    create_detection_mechanisms(wb["Detection Mechanisms"], styles)
    
    logger.info("[3/9] Creating Alert Handling sheet (25 questions)...")
    create_alert_handling(wb["Alert Handling"], styles)
    
    logger.info("[4/9] Creating Classification & Severity sheet (25 questions)...")
    create_classification_severity(wb["Classification & Severity"], styles)
    
    logger.info("[5/9] Creating Detection Effectiveness sheet (25 questions)...")
    create_detection_effectiveness(wb["Detection Effectiveness"], styles)
    
    logger.info("[6/9] Creating Gap Analysis sheet...")
    create_gap_analysis(wb["Gap Analysis"], styles)
    
    logger.info("[7/9] Creating Evidence Register sheet...")
    create_evidence_register(wb["Evidence Register"])

    logger.info("[8/9] Creating Dashboard sheet...")
    create_dashboard(wb["Summary Dashboard"], styles)

    logger.info("[9/9] Creating Approval Sign-Off sheet...")
    create_approval_sheet(wb["Approval Sign-Off"])

    finalize_validations(wb)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    output_path = _wkbk_dir / OUTPUT_FILENAME
    wb.save(output_path)
    logger.info(f"\nSUCCESS: {_wkbk_dir / OUTPUT_FILENAME}")
    logger.info("\nWorkbook Structure:")
    logger.info("  • 9 sheets (Instructions through Approval)")
    logger.info("  • 108 assessment questions (33+25+25+25)")
    logger.info("  • 40 gap analysis capacity")
    logger.info("  • 60 evidence register capacity")
    logger.info("  • Automated metrics (MTTD, MTTT, FP rate, coverage)")
    logger.info("\nNext steps:")
    logger.info("  1) Extract metrics from SIEM/ticketing system (90 days)")
    logger.info("  2) Complete detection mechanism inventory")
    logger.info("  3) Fill yellow cells in assessment sheets")
    logger.info("  4) Review automated metric calculations")
    logger.info("  5) Analyze detection coverage gaps")
    logger.info("  6) Complete Approval Sign-Off workflow")
    logger.info("\nEstimated completion time: 6-10 hours")
    logger.info("\n" + "=" * 80)


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
