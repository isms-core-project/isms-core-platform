#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.24-28.S5 - Learning & Continuous Improvement Excel Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.27: Learning from Information Security Incidents
Assessment Domain 5 of 5: Learning & Continuous Improvement

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMISATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific incident management infrastructure, technology stack,
and assessment requirements.

Key customisation areas:
1. Incident classification levels and severity criteria (match your organisation)
2. Detection source categories and integration points (adapt to your toolset)
3. Response team roles and escalation thresholds
4. Evidence collection requirements and chain-of-custody procedures
5. Lessons-learned process and improvement tracking mechanisms

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.27 Learning from Information Security Incidents Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
incident management controls and compliance requirements.

**Purpose:**
Enables systematic assessment of Learning & Continuous Improvement under ISO 27001:2022 Controls A.5.24-A.5.28. Supports evidence-based evaluation of incident lifecycle capabilities, response effectiveness, and continuous improvement processes.

**Assessment Scope:**
- Incident detection mechanism coverage and response time metrics
- Classification accuracy and severity assignment consistency
- Response capability completeness and team readiness
- Forensic evidence collection and preservation procedures
- Stakeholder communication and regulatory notification compliance
- Lessons-learned implementation and effectiveness tracking
- Evidence collection for incident management and regulatory audits

**Generated Workbook Structure:**
1. Instructions & Legend - Assessment guidance and scoring methodology
2. [Data sheets] - Assessment data input sheets
6. Summary Dashboard - Compliance overview and key metrics
7. Evidence Register - Audit evidence tracking
8. Approval Sign-Off - Stakeholder review and approval workflow

**Key Features:**
- Data validation with standardised dropdown lists
- Conditional formatting for visual compliance status
- Automated compliance scoring and gap identification
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow

**Integration:**
This assessment is one of 5 domains covering Learning from Information Security Incidents controls.
Results feed into the Summary Dashboard for executive oversight.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a524_28_s5_learning_improvement.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a524_28_s5_learning_improvement.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a524_28_s5_learning_improvement.py --date 20250115

Output:
    File: ISMS-IMP-A.5.24-28.S5_Learning_&_Continuous_Improvement_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Review the Instructions & Legend sheet for assessment guidance
    2. Populate the assessment data sheets with your organisation's information
    3. Complete all required fields marked with yellow (FFFFCC) highlighting
    4. Review automated compliance calculations in the Summary Dashboard
    5. Document gaps and assign remediation owners in Gap Analysis sheets
    6. Collect and link audit evidence in the Evidence Register
    7. Obtain stakeholder sign-off via the Approval Sign-Off sheet
    8. Review Summary Dashboard metrics and finalise compliance reporting

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.27
Assessment Domain:    5 of 5 (Learning & Continuous Improvement)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.27: Learning from Information Security Incidents Policy (Governance)
    - ISMS-IMP-A.5.24-28.S1: Incident Management Framework (Domain 1)
    - ISMS-IMP-A.5.24-28.S2: Detection & Classification (Domain 2)
    - ISMS-IMP-A.5.24-28.S3: Response Capabilities (Domain 3)
    - ISMS-IMP-A.5.24-28.S4: Forensic Evidence (Domain 4)
    - ISMS-IMP-A.5.24-28.S5: Learning & Continuous Improvement (Domain 5)

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.24-28.S5 specification
    - Supports compliance tracking and gap identification
    - Supports integrated Summary Dashboard reporting

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Ensure all fields are completed accurately and evidence is properly linked.

**Data Protection:**
Assessment workbooks may contain sensitive incident management details. Handle
in accordance with your organisation's data classification policies.

**Maintenance:**
Review incident classification criteria, response procedures, and lessons-learned outcomes annually or after significant incidents, major infrastructure changes, or regulatory updates.

**Quality Assurance:**
Have technical SMEs validate assessments before using results
for compliance reporting or management decisions.

================================================================================
"""

from datetime import datetime
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("Error: openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging
import sys

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.24-28.S5"
WORKBOOK_NAME = "Learning & Continuous Improvement"
CONTROL_ID = "A.5.27"
CONTROL_NAME = "Learning from Information Security Incidents"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # For display (Swiss format)
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"
_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point

# ============================================================================
# SECTION 1: WORKBOOK CREATION & STYLE DEFINITIONS
# ============================================================================

def create_workbook() -> Workbook:
    """Create workbook with all required sheets matching IMP specification."""
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb.active)

    # Sheet structure per ISMS-IMP-A.5.24-28.S5 specification
    sheets = [
        "Instructions & Legend",
        "PIR Process",
        "Root Cause Analysis",
        "Lessons Learned",
        "Control Improvements",
        "Trend Analysis",
        "Gap Analysis",
        "Evidence Register",
        "Summary Dashboard",
        "Approval Sign-Off",
    ]
    for name in sheets:
        wb.create_sheet(title=name)

    return wb


def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    styles = {
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "column_header": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        "section_header": {
            "font": Font(name="Calibri", size=11, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
        },
        "input": {
            "font": Font(name="Calibri", size=10),
            "fill": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "formula": {
            "font": Font(name="Calibri", size=10),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "normal": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        },
        "compliant": {
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "partial": {
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "non_compliant": {
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "border": border_thin,
    }
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict, border=None):
    """Apply style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = style_dict["font"]
    if "fill" in style_dict:
        cell.fill = style_dict["fill"]
    if "alignment" in style_dict:
        cell.alignment = style_dict["alignment"]
    if border:
        cell.border = border


# ============================================================================
# SECTION 2: INSTRUCTIONS & LEGEND SHEET
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Complete PIR Process sheet — review all PIRs from last 12 months.', '2. Complete Root Cause Analysis sheet — assess RCA quality for Critical/High incidents.', '3. Complete Lessons Learned sheet — verify documentation and distribution.', '4. Complete Control Improvements sheet — track remediation actions.', '5. Complete Trend Analysis sheet — verify KPIs and reporting.', '6. Review Gap Analysis.', '7. Link evidence in Evidence Register.', '8. Review Summary Dashboard for overall compliance.', '9. Complete Approval Sign-Off workflow.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_pir_process_sheet(ws, styles):
    """Create the PIR Process assessment sheet."""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12

    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = "PIR PROCESS ASSESSMENT"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 35

    ws['A2'] = f"{CONTROL_REF} | Post-Incident Review timeliness, quality scoring, and SLA compliance"
    ws['A2'].font = Font(name="Calibri", size=10, italic=True)
    ws.merge_cells('A2:L2')

    ws.freeze_panes = "A4"

    # Section A header
    ws.merge_cells('A3:L3')
    ws['A3'] = "Section A: PIR Inventory (Last 12 Months)"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    # Column headers
    headers = [
        "Incident ID", "Incident Date", "Severity", "Resolution Date",
        "PIR Status", "PIR Completion Date", "SLA Days", "Actual Days",
        "SLA Met", "Participants Met", "Quality Score", "Evidence Ref"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["column_header"], styles["border"])

    # Sample row (F2F2F2 grey) at row 6 — 1 example entry to guide assessors
    _f2f2f2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_font = Font(name="Calibri", size=10, italic=True, color="808080")
    _sample_data = ["INC-001", "01.01.2026", "High", "04.01.2026", "Completed",
                    "08.01.2026", "5", "7", "Yes", "Yes", "4", "EV-001"]
    for col_idx, val in enumerate(_sample_data, start=1):
        cell = ws.cell(row=6, column=col_idx, value=val)
        cell.fill = _f2f2f2
        cell.font = _sample_font
        cell.border = styles["border"]

    # Data rows (50 empty FFFFCC rows for PIR inventory)
    for row in range(7, 57):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col, value="")
            apply_style(cell, styles["input"], styles["border"])

    # Data validation for Severity
    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(severity_dv)
    severity_dv.add(f'C7:C56')

    # Data validation for PIR_Status
    status_dv = DataValidation(type="list", formula1='"Completed,Overdue,Pending,Not_Required"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f'E7:E56')

    # Data validation for Yes/No fields
    yesno_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yesno_dv)
    yesno_dv.add(f'I7:I56')
    yesno_dv.add(f'J7:J56')

    # Data validation for Quality Score
    quality_dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(quality_dv)
    quality_dv.add(f'K7:K56')

    # Section B: Quality Scoring Criteria (reference)
    row = 58
    ws.merge_cells(f'A{row}:L{row}')
    ws[f'A{row}'] = "Section B: PIR Quality Scoring Criteria (Reference)"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    criteria = [
        ("5 - Excellent", "Complete timeline, full impact assessment, all action items with owners/dates, root cause referenced, distributed"),
        ("4 - Good", "Timeline and impact documented, most action items defined, minor gaps in distribution"),
        ("3 - Adequate", "Basic incident summary, some action items defined, distribution incomplete"),
        ("2 - Poor", "Minimal documentation, few action items, no distribution"),
        ("1 - Non-existent", "PIR marked complete but no substantive report exists"),
    ]

    for score, desc in criteria:
        row += 1
        ws[f'A{row}'] = score
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws.merge_cells(f'B{row}:L{row}')
        ws[f'B{row}'] = desc
        apply_style(ws[f'B{row}'], styles["normal"], styles["border"])

    # Section C: Summary
    row += 2
    ws.merge_cells(f'A{row}:L{row}')
    ws[f'A{row}'] = "Section C: PIR Process Summary"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    summary_items = [
        ("Total incidents in period", '=COUNTA(A7:A56)'),
        ("PIRs completed", '=COUNTIF(E7:E56,"Completed")'),
        ("PIRs overdue", '=COUNTIF(E7:E56,"Overdue")'),
        ("SLA compliance rate (%)", '=IF(B' + str(row+2) + '>0,COUNTIF(I7:I56,"Yes")/B' + str(row+2) + '*100,0)'),
        ("Average quality score", '=AVERAGE(K7:K56)'),
        ("Participant requirement met (%)", '=IF(B' + str(row+2) + '>0,COUNTIF(J7:J56,"Yes")/B' + str(row+2) + '*100,0)'),
    ]

    for label, formula in summary_items:
        row += 1
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'D{row}'] = formula
        apply_style(ws[f'D{row}'], styles["formula"], styles["border"])


# ============================================================================
# SECTION 4: ROOT CAUSE ANALYSIS SHEET
# ============================================================================

def create_rca_sheet(ws, styles):
    """Create the Root Cause Analysis assessment sheet."""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12

    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = "ROOT CAUSE ANALYSIS ASSESSMENT"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 35

    ws['A2'] = f"{CONTROL_REF} | RCA inventory, depth scoring, methodology, and recurring cause identification"
    ws['A2'].font = Font(name="Calibri", size=10, italic=True)
    ws.merge_cells('A2:I2')

    ws.freeze_panes = "A4"

    # Section A header
    ws.merge_cells('A3:I3')
    ws['A3'] = "Section A: RCA Inventory (Critical and High Incidents)"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    # Column headers
    headers = [
        "Incident ID", "Severity", "RCA Status", "RCA Date",
        "Methodology", "Root Cause Summary", "Depth Score", "Recurring", "Evidence Ref"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["column_header"], styles["border"])

    # Sample row (F2F2F2 grey) at row 6 — 1 example entry to guide assessors
    _f2f2f2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_font = Font(name="Calibri", size=10, italic=True, color="808080")
    _sample_rca = ["INC-001", "Critical", "Completed", "05.01.2026",
                   "5_Whys", "Misconfigured firewall rule allowed lateral movement — sample entry",
                   "3 - Systemic", "No", "EV-001"]
    for col_idx, val in enumerate(_sample_rca, start=1):
        cell = ws.cell(row=6, column=col_idx, value=val)
        cell.fill = _f2f2f2
        cell.font = _sample_font
        cell.border = styles["border"]

    # Data rows (50 rows for Critical/High incidents)
    for row in range(7, 57):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col, value="")
            apply_style(cell, styles["input"], styles["border"])

    # Data validation for Severity
    severity_dv = DataValidation(type="list", formula1='"Critical,High"', allow_blank=True)
    ws.add_data_validation(severity_dv)
    severity_dv.add(f'B7:B56')

    # Data validation for RCA_Status
    status_dv = DataValidation(type="list", formula1='"Completed,In_Progress,Not_Performed"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f'C7:C56')

    # Data validation for Methodology
    method_dv = DataValidation(type="list", formula1='"5_Whys,Fishbone,Fault_Tree,Timeline,Combined,Other"', allow_blank=True)
    ws.add_data_validation(method_dv)
    method_dv.add(f'E7:E56')

    # Data validation for Depth Score
    depth_dv = DataValidation(type="list", formula1='"1 - Technical,2 - Procedural,3 - Systemic"', allow_blank=True)
    ws.add_data_validation(depth_dv)
    depth_dv.add(f'G7:G56')

    # Data validation for Recurring
    yesno_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yesno_dv)
    yesno_dv.add(f'H7:H56')

    # Section B: RCA Quality Assessment
    row = 59
    ws.merge_cells(f'A{row}:I{row}')
    ws[f'A{row}'] = "Section B: RCA Quality Questions"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    quality_questions = [
        "Does RCA go beyond immediate technical cause to identify process/governance gaps?",
        "Was the RCA methodology appropriate for the incident complexity?",
        "Are preventive recommendations specific and actionable?",
        "Has the RCA been reviewed by someone other than the incident owner?",
    ]

    for q in quality_questions:
        row += 1
        ws[f'A{row}'] = q
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'H{row}'] = ""
        apply_style(ws[f'H{row}'], styles["input"], styles["border"])

    # Data validation for quality questions
    yesno_dv2 = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    ws.add_data_validation(yesno_dv2)
    yesno_dv2.add(f'H60:H63')

    # Section C: Summary
    row += 3
    ws.merge_cells(f'A{row}:I{row}')
    ws[f'A{row}'] = "Section C: RCA Summary"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    summary_items = [
        ("Critical/High incidents in period", '=COUNTA(A7:A56)'),
        ("RCAs completed", '=COUNTIF(C7:C56,"Completed")'),
        ("RCAs not performed", '=COUNTIF(C7:C56,"Not_Performed")'),
        ("Systemic depth RCAs", '=COUNTIF(G7:G56,"3 - Systemic")'),
        ("Recurring root causes", '=COUNTIF(H7:H56,"Yes")'),
    ]

    for label, formula in summary_items:
        row += 1
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'D{row}'] = formula
        apply_style(ws[f'D{row}'], styles["formula"], styles["border"])


# ============================================================================
# SECTION 5: LESSONS LEARNED SHEET
# ============================================================================

def create_lessons_learned_sheet(ws, styles):
    """Create the Lessons Learned assessment sheet."""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "LESSONS LEARNED ASSESSMENT"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 35

    ws['A2'] = f"{CONTROL_REF} | Lessons learned log, knowledge base inventory, and governance verification"
    ws['A2'].font = Font(name="Calibri", size=10, italic=True)
    ws.merge_cells('A2:H2')

    ws.freeze_panes = "A4"

    # Section A header
    ws.merge_cells('A3:H3')
    ws['A3'] = "Section A: Lessons Learned Log (Per PIR)"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    # Column headers
    headers = [
        "PIR ID", "LL Entry Date", "Lesson Summary", "Distribution Date",
        "SLA Met", "Playbook Update", "Playbook Date", "Evidence Ref"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["column_header"], styles["border"])

    # Sample row (F2F2F2 grey) at row 6 — 1 example entry to guide assessors
    _f2f2f2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_font = Font(name="Calibri", size=10, italic=True, color="808080")
    _sample_ll = ["PIR-001", "08.01.2026", "Firewall rules must be reviewed quarterly — sample entry",
                  "15.01.2026", "Yes", "Yes", "20.01.2026", "EV-001"]
    for col_idx, val in enumerate(_sample_ll, start=1):
        cell = ws.cell(row=6, column=col_idx, value=val)
        cell.fill = _f2f2f2
        cell.font = _sample_font
        cell.border = styles["border"]

    # Data rows (50 rows)
    for row in range(7, 57):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col, value="")
            apply_style(cell, styles["input"], styles["border"])

    # Data validation for Yes/No fields
    yesno_dv = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=True)
    ws.add_data_validation(yesno_dv)
    yesno_dv.add(f'E7:E56')
    yesno_dv.add(f'F7:F56')

    # Section B: Knowledge Base Inventory
    row = 59
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "Section B: Knowledge Base Inventory"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    kb_headers = ["KB ID", "Title", "Type", "Last Updated", "Status", "Owner", "Evidence Ref", ""]
    row += 2
    for col, header in enumerate(kb_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, styles["column_header"], styles["border"])

    # Data rows for KB (15 rows)
    for r in range(row + 1, row + 16):
        for col in range(1, 8):
            cell = ws.cell(row=r, column=col, value="")
            apply_style(cell, styles["input"], styles["border"])

    # Data validation for KB Type
    type_dv = DataValidation(type="list", formula1='"Playbook,Procedure,Template,Reference,Training"', allow_blank=True)
    ws.add_data_validation(type_dv)
    type_dv.add(f'C{row+1}:C{row+15}')

    # Data validation for KB Status
    status_dv = DataValidation(type="list", formula1='"Current,Outdated,Missing,Under_Review"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f'E{row+1}:E{row+15}')

    # Section C: Summary
    row += 18
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "Section C: Lessons Learned Summary"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    summary_items = [
        ("Total lessons learned entries", '=COUNTA(A7:A56)'),
        ("Distribution SLA met (%)", '=IF(COUNTA(A7:A56)>0,COUNTIF(E7:E56,"Yes")/COUNTA(A7:A56)*100,0)'),
        ("Playbook updates completed", '=COUNTIF(F7:F56,"Yes")'),
        ("KB items current", '=COUNTIF(E62:E76,"Current")'),
        ("KB items outdated/missing", '=COUNTIFS(E62:E76,"<>Current",E62:E76,"<>")'),
    ]

    for label, formula in summary_items:
        row += 1
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'D{row}'] = formula
        apply_style(ws[f'D{row}'], styles["formula"], styles["border"])


# ============================================================================
# SECTION 6: CONTROL IMPROVEMENTS SHEET
# ============================================================================

def create_control_improvements_sheet(ws, styles):
    """Create the Control Improvements assessment sheet."""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 12

    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = "CONTROL IMPROVEMENTS ASSESSMENT"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 35

    ws['A2'] = f"{CONTROL_REF} | Remediation action register, closure tracking, and overdue escalation"
    ws['A2'].font = Font(name="Calibri", size=10, italic=True)
    ws.merge_cells('A2:K2')

    ws.freeze_panes = "A4"

    # Section A header
    ws.merge_cells('A3:K3')
    ws['A3'] = "Section A: Remediation Action Register"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    # Column headers
    headers = [
        "Action ID", "Source Incident", "Action Description", "Priority",
        "Owner", "Target Date", "Status", "Completion Date", "Verified By",
        "Escalated", "Evidence Ref"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["column_header"], styles["border"])

    # Sample row (F2F2F2 grey) at row 6 — 1 example entry to guide assessors
    _f2f2f2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_font = Font(name="Calibri", size=10, italic=True, color="808080")
    _sample_ci = ["ACT-001", "INC-001", "Update firewall rules and review quarterly — sample entry",
                  "High", "IT Manager", "28.02.2026", "Open", "", "", "No", "EV-001"]
    for col_idx, val in enumerate(_sample_ci, start=1):
        cell = ws.cell(row=6, column=col_idx, value=val)
        cell.fill = _f2f2f2
        cell.font = _sample_font
        cell.border = styles["border"]

    # Data rows (50 rows)
    for row in range(7, 57):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col, value="")
            apply_style(cell, styles["input"], styles["border"])

    # Data validation for Priority
    priority_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(priority_dv)
    priority_dv.add(f'D7:D56')

    # Data validation for Status
    status_dv = DataValidation(type="list", formula1='"Open,In_Progress,Completed,Blocked,Cancelled"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f'G7:G56')

    # Data validation for Yes/No
    yesno_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yesno_dv)
    yesno_dv.add(f'J7:J56')

    # Section B: Summary
    row = 59
    ws.merge_cells(f'A{row}:K{row}')
    ws[f'A{row}'] = "Section B: Action Summary Metrics"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    summary_items = [
        ("Total actions", '=COUNTA(A7:A56)'),
        ("Actions completed", '=COUNTIF(G7:G56,"Completed")'),
        ("Actions open", '=COUNTIF(G7:G56,"Open")'),
        ("Actions blocked", '=COUNTIF(G7:G56,"Blocked")'),
        ("Closure rate (%)", '=IF(COUNTA(A7:A56)>0,COUNTIF(G7:G56,"Completed")/COUNTA(A7:A56)*100,0)'),
        ("Critical actions open", '=COUNTIFS(D7:D56,"Critical",G7:G56,"<>Completed")'),
        ("Overdue actions", '=COUNTIFS(F7:F56,"<"&TODAY(),G7:G56,"<>Completed")'),
    ]

    for label, formula in summary_items:
        row += 1
        ws[f'A{row}'] = label
        apply_style(ws[f'A{row}'], styles["normal"], styles["border"])
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'D{row}'] = formula
        apply_style(ws[f'D{row}'], styles["formula"], styles["border"])


# ============================================================================
# SECTION 7: TREND ANALYSIS SHEET
# ============================================================================

def create_trend_analysis_sheet(ws, styles):
    """Create the Trend Analysis assessment sheet."""
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "TREND ANALYSIS ASSESSMENT"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 35

    ws['A2'] = f"{CONTROL_REF} | KPI definitions, current values, trend tracking, and reporting cadence"
    ws['A2'].font = Font(name="Calibri", size=10, italic=True)
    ws.merge_cells('A2:H2')

    ws.freeze_panes = "A4"

    # Section A header
    ws.merge_cells('A3:H3')
    ws['A3'] = "Section A: KPI Definitions and Current Values"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    # KPI headers
    kpi_headers = ["KPI", "Metric Type", "Target", "Current", "Status", "Trend", "Period", "Evidence Ref"]
    for col, header in enumerate(kpi_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["column_header"], styles["border"])

    # Pre-defined KPIs
    kpis = [
        ("Incident Recurrence Rate", "Percentage", "<=15%", "", "", "", "", ""),
        ("PIR SLA Compliance Rate", "Percentage", ">=95%", "", "", "", "", ""),
        ("Lessons Learned Distribution SLA", "Percentage", ">=90%", "", "", "", "", ""),
        ("Remediation Closure Rate", "Percentage", ">=85%", "", "", "", "", ""),
        ("Average PIR Quality Score", "Score (1-5)", ">=4.0", "", "", "", "", ""),
        ("RCA Completion Rate (Crit/High)", "Percentage", "100%", "", "", "", "", ""),
        ("MTTD (Critical Incidents)", "Hours", "<=1", "", "", "", "", ""),
        ("MTTR (Critical Incidents)", "Hours", "<=4", "", "", "", "", ""),
    ]

    for row_num, kpi_data in enumerate(kpis, 6):
        for col, value in enumerate(kpi_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            if col <= 3:
                apply_style(cell, styles["normal"], styles["border"])
            else:
                apply_style(cell, styles["input"], styles["border"])

    # Data validation for Status
    status_dv = DataValidation(type="list", formula1='"Yes,No,At_Risk"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f'E6:E13')

    # Data validation for Trend
    trend_dv = DataValidation(type="list", formula1='"Improving,Stable,Degrading"', allow_blank=True)
    ws.add_data_validation(trend_dv)
    trend_dv.add(f'F6:F13')

    # Section B: Reporting Cadence Verification
    row = 16
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "Section B: Reporting Cadence Verification"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])

    report_headers = ["Report Type", "Frequency", "Last Produced", "On Schedule", "Recipients Documented", "Evidence Ref", "", ""]
    row += 2
    for col, header in enumerate(report_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, styles["column_header"], styles["border"])

    reports = [
        ("Weekly CSIRT Status", "Weekly", "", "", "", ""),
        ("Monthly Incident Summary", "Monthly", "", "", "", ""),
        ("Quarterly Trend Report", "Quarterly", "", "", "", ""),
        ("Annual Executive Report", "Annual", "", "", "", ""),
    ]

    for row_num, report_data in enumerate(reports, row + 1):
        for col, value in enumerate(report_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            if col <= 2:
                apply_style(cell, styles["normal"], styles["border"])
            else:
                apply_style(cell, styles["input"], styles["border"])

    # Data validation for Yes/No
    yesno_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yesno_dv)
    yesno_dv.add(f'D{row+1}:E{row+4}')


# ============================================================================
# SECTION 8: GAP ANALYSIS SHEET
# ============================================================================

def create_gap_analysis_sheet(ws, styles):
    """Create the Gap Analysis sheet."""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 12

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "GAP ANALYSIS"
    apply_style(ws['A1'], styles["header"], styles["border"])
    ws.row_dimensions[1].height = 35

    ws['A2'] = f"{CONTROL_REF} | Consolidated gap register for Learning & Continuous Improvement domain"
    ws['A2'].font = Font(name="Calibri", size=10, italic=True)
    ws.merge_cells('A2:H2')

    ws.freeze_panes = "A4"

    ws.merge_cells('A3:H3')
    ws['A3'] = "Consolidated Gap Register"
    apply_style(ws['A3'], styles["section_header"], styles["border"])

    # Column headers
    headers = [
        "Gap ID", "Source Sheet", "Gap Description", "Severity",
        "Owner", "Target Date", "Remediation Plan", "Status"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_style(cell, styles["column_header"], styles["border"])

    # Sample row (F2F2F2 grey) at row 6 — 1 example entry to guide assessors
    _f2f2f2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _sample_font = Font(name="Calibri", size=10, italic=True, color="808080")
    _sample_gap = ["GAP-001", "PIR Process", "PIR completion SLA not consistently met — sample entry",
                   "High", "CSIRT Manager", "31.03.2026", "Establish automated SLA tracking", "Open"]
    for col_idx, val in enumerate(_sample_gap, start=1):
        cell = ws.cell(row=6, column=col_idx, value=val)
        cell.fill = _f2f2f2
        cell.font = _sample_font
        cell.border = styles["border"]

    # Data rows (50 rows)
    for row in range(7, 57):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col, value="")
            apply_style(cell, styles["input"], styles["border"])

    # Data validation for Source_Sheet
    source_dv = DataValidation(type="list", formula1='"PIR Process,Root Cause Analysis,Lessons Learned,Control Improvements,Trend Analysis"', allow_blank=True)
    ws.add_data_validation(source_dv)
    source_dv.add(f'B7:B56')

    # Data validation for Severity
    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(severity_dv)
    severity_dv.add(f'D7:D56')

    # Data validation for Status
    status_dv = DataValidation(type="list", formula1='"Open,In_Progress,Closed,Accepted"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f'H7:H56')

    # Summary row
    row = 59
    ws[f'A{row}'] = "Total Gaps:"
    apply_style(ws[f'A{row}'], styles["section_header"], styles["border"])
    ws[f'B{row}'] = '=COUNTA(A7:A56)'
    apply_style(ws[f'B{row}'], styles["formula"], styles["border"])
    ws[f'C{row}'] = "Critical:"
    ws[f'D{row}'] = '=COUNTIF(D7:D56,"Critical")'
    apply_style(ws[f'D{row}'], styles["formula"], styles["border"])
    ws[f'E{row}'] = "High:"
    ws[f'F{row}'] = '=COUNTIF(D7:D56,"High")'
    apply_style(ws[f'F{row}'], styles["formula"], styles["border"])


# ============================================================================
# SECTION 9: EVIDENCE REGISTER SHEET
# ============================================================================

def create_evidence_register(ws):
    """Create the standard Evidence Register sheet (gold standard)."""
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _input = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = _border

    for c in range(1, 9):
        ws.cell(row=3, column=c).border = _border

    columns = [
        ("Evidence ID", 14), ("Evidence Type", 20), ("Description", 45),
        ("Related Control / Section", 28), ("Collection Date (DD.MM.YYYY)", 22),
        ("Storage Location / Reference", 38), ("Collected By", 22), ("Status", 14),
    ]
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    sample_data = ["EV-001", "Document", "Sample evidence entry — replace with actual evidence",
                   "All Controls", "01.01.2026", "SharePoint/ISMS/Evidence/", "ISMS Team", "Active"]
    for col_idx, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
        cell.fill = _grey_sample
        cell.border = _border

    dv_status = DataValidation(
        type="list",
        formula1='"Active,Archived,Superseded,Pending Review"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(6, 106):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = _input
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"


# ============================================================================
# SECTION 10: SUMMARY DASHBOARD SHEET
# ============================================================================

def create_summary_dashboard_sheet(ws, styles):
    """Summary Dashboard — Gold Standard TABLE 1/2/3."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{WORKBOOK_NAME} — SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=1, column=_c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: CONTROL_REF subtitle
    ws.merge_cells("A2:G2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=2, column=_c).border = border

    # TABLE 1 banner (row 3)
    ws.merge_cells("A3:G3")
    ws["A3"] = "COMPLIANCE ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=3, column=_c).border = border

    # TABLE 1 column headers (row 4)
    _t1_headers = ['Assessment Area', 'Records', 'Completed', 'In Progress', 'Not Done', 'Target', 'Completion %']
    for _c, _h in enumerate(_t1_headers, 1):
        _cell = ws.cell(row=4, column=_c, value=_h)
        _cell.font = Font(bold=True, size=10, color="000000", name="Calibri")
        _cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _cell.border = border
    ws.row_dimensions[4].height = 30

    # TABLE 1 area rows — S5 incident-tracking columns (per-sheet status col)
    _s5_areas = [
        # (label, col B formula, col C formula, col D formula, col E formula, target)
        ("PIR Process",
         "=COUNTA('PIR Process'!A6:A100)",
         "=COUNTIF('PIR Process'!E5:E100,\"Completed\")",
         "=COUNTIF('PIR Process'!E5:E100,\"Pending\")",
         "=COUNTIF('PIR Process'!E5:E100,\"Overdue\")",
         "95%"),
        ("Root Cause Analysis",
         "=COUNTA('Root Cause Analysis'!A6:A100)",
         "=COUNTIF('Root Cause Analysis'!C5:C100,\"Completed\")",
         "=COUNTIF('Root Cause Analysis'!C5:C100,\"In_Progress\")",
         "=COUNTIF('Root Cause Analysis'!C5:C100,\"Not_Performed\")",
         "90%"),
        ("Lessons Learned",
         "=COUNTA('Lessons Learned'!A6:A100)",
         "=COUNTIF('Lessons Learned'!F5:F100,\"Yes\")",
         "=COUNTIF('Lessons Learned'!F5:F100,\"N/A\")",
         "=COUNTIF('Lessons Learned'!F5:F100,\"No\")",
         "85%"),
        ("Control Improvements",
         "=COUNTA('Control Improvements'!A6:A100)",
         "=COUNTIF('Control Improvements'!G5:G100,\"Completed\")",
         "=COUNTIF('Control Improvements'!G5:G100,\"In_Progress\")+COUNTIF('Control Improvements'!G5:G100,\"Open\")",
         "=COUNTIF('Control Improvements'!G5:G100,\"Blocked\")+COUNTIF('Control Improvements'!G5:G100,\"Cancelled\")",
         "85%"),
        ("Trend Analysis",
         "=COUNTA('Trend Analysis'!A6:A100)",
         "=COUNTIF('Trend Analysis'!E5:E100,\"Yes\")",
         "=COUNTIF('Trend Analysis'!E5:E100,\"At_Risk\")",
         "=COUNTIF('Trend Analysis'!E5:E100,\"No\")",
         "80%"),
    ]
    for _i, (_label, _b, _c, _d, _e, _target) in enumerate(_s5_areas, start=5):
        _r = _i
        ws.cell(row=_r, column=1).value = _label
        ws.cell(row=_r, column=2).value = _b
        ws.cell(row=_r, column=3).value = _c
        ws.cell(row=_r, column=4).value = _d
        ws.cell(row=_r, column=5).value = _e
        ws.cell(row=_r, column=6).value = _target
        ws.cell(row=_r, column=7).value = f"=IFERROR(C{_r}/(C{_r}+D{_r}+E{_r})*100,0)"
        ws.cell(row=_r, column=7).number_format = "0.0"
        for _c_idx in range(1, 8):
            _cell = ws.cell(row=_r, column=_c_idx)
            _cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            _cell.border = border
            _cell.alignment = Alignment(horizontal="left", vertical="center")

    # TABLE 1 TOTAL row (row 10)
    ws.cell(row=10, column=1).value = "TOTAL"
    ws.cell(row=10, column=2).value = "=SUM(B5:B9)"
    ws.cell(row=10, column=3).value = "=SUM(C5:C9)"
    ws.cell(row=10, column=4).value = "=SUM(D5:D9)"
    ws.cell(row=10, column=5).value = "=SUM(E5:E9)"
    ws.cell(row=10, column=6).value = "—"
    ws.cell(row=10, column=7).value = "=IFERROR(AVERAGE(G5:G9),0)"
    ws.cell(row=10, column=7).number_format = "0.0"
    for _c in range(1, 8):
        _cell = ws.cell(row=10, column=_c)
        _cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _cell.font = Font(bold=True, name="Calibri")
        _cell.border = border
        _cell.alignment = Alignment(horizontal="left", vertical="center")
    _total_row = 10

    # TABLE 2
    _t2_row = _total_row + 2
    ws.merge_cells(f"A{_t2_row}:G{_t2_row}")
    ws[f"A{_t2_row}"] = 'LEARNING PROGRAMME METRICS'
    ws[f"A{_t2_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{_t2_row}"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws[f"A{_t2_row}"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=_t2_row, column=_c).border = border
    _t2_row += 1
    # TABLE 2 col headers
    _t2_hdrs = ["Metric", "Value", "Target"]
    for _c, _h in enumerate(_t2_hdrs, 1):
        _cell = ws.cell(row=_t2_row, column=_c, value=_h)
        _cell.font = Font(bold=True, size=10, color="000000", name="Calibri")
        _cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _cell.alignment = Alignment(horizontal="center", vertical="center")
        _cell.border = border
    _t2_row += 1
    # TABLE 2 metric rows
    _t2_metrics = [('PIR Completion Rate (%)', "='PIR Process'!D72", '≥95%'), ('RCA Completion Rate (%)', "='Root Cause Analysis'!D70", '≥90%'), ('Lessons Learned Action Rate (%)', "='Lessons Learned'!D82", '≥85%'), ('Control Improvement Closure Rate (%)', "='Control Improvements'!D63", '≥80%')]
    for _label, _formula, _target in _t2_metrics:
        ws.cell(row=_t2_row, column=1).value = _label
        ws.cell(row=_t2_row, column=2).value = _formula
        ws.cell(row=_t2_row, column=3).value = _target
        for _c in range(1, 4):
            _cell = ws.cell(row=_t2_row, column=_c)
            _cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            _cell.border = border
            _cell.alignment = Alignment(horizontal="left", vertical="center")
        _t2_row += 1
    _t3_start = _t2_row

    # TABLE 3
    _t3_row = _t3_start + 1
    ws.merge_cells(f"A{_t3_row}:G{_t3_row}")
    ws[f"A{_t3_row}"] = 'CRITICAL FINDINGS'
    ws[f"A{_t3_row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{_t3_row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws[f"A{_t3_row}"].alignment = Alignment(horizontal="center", vertical="center")
    for _c in range(1, 8):
        ws.cell(row=_t3_row, column=_c).border = border
    _t3_row += 1
    # TABLE 3 col headers
    _t3_hdrs = ["Critical Finding", "Status", "Severity"]
    for _c, _h in enumerate(_t3_hdrs, 1):
        _cell = ws.cell(row=_t3_row, column=_c, value=_h)
        _cell.font = Font(bold=True, size=10, color="000000", name="Calibri")
        _cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _cell.alignment = Alignment(horizontal="center", vertical="center")
        _cell.border = border
    _t3_row += 1
    _t3_findings = [('PIR completion <70%', '=IF(G5<70,"[!] Below Target","[OK]")', 'Critical'), ('RCA completion <70%', '=IF(G6<70,"[!] Below Target","[OK]")', 'Critical'), ('Lessons learned actioned <70%', '=IF(G7<70,"[!] Below Target","[OK]")', 'High'), ('Control improvements <70%', '=IF(G8<70,"[!] Below Target","[OK]")', 'High'), ('Trend analysis KPIs <70%', '=IF(G9<70,"[!] Below Target","[OK]")', 'High'), ('Overall learning rate <70%', '=IF(G10<70,"[!] Below Target","[OK]")', 'Critical')]
    _t3_sev_fills = {'Critical': 'FFC7CE', 'High': 'FFEB9C', 'Low': 'C6EFCE'}
    for _label, _formula, _severity in _t3_findings:
        _fill_color = _t3_sev_fills.get(_severity, "FFFFCC")
        ws.cell(row=_t3_row, column=1).value = _label
        ws.cell(row=_t3_row, column=2).value = _formula
        ws.cell(row=_t3_row, column=3).value = _severity
        for _c in range(1, 4):
            _cell = ws.cell(row=_t3_row, column=_c)
            _cell.fill = PatternFill(start_color=_fill_color, end_color=_fill_color, fill_type="solid")
            _cell.border = border
            _cell.alignment = Alignment(horizontal="left", vertical="center")
        _t3_row += 1

    # Column widths + freeze
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A4"


def create_approval_sheet(ws):
    """Create the Approval Sign-Off sheet — Gold Standard."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Summary fields (rows 4-8)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G9),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    # GS-AS-015: Overall Completion Rate — must reference Summary Dashboard
    ws["B6"].value = "=IFERROR(AVERAGE('Summary Dashboard'!G5:G9),\"\")"
    ws["B6"].number_format = "0.0%"

    # Assessment Status dropdown (row 7)
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # 3 Approver sections (start at row 11)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"


# ============================================================================
# SECTION 12: FINALISE & MAIN EXECUTION
# ============================================================================

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        ws.data_validations.dataValidation = [
            dv for dv in list(ws.data_validations.dataValidation)
            if dv.sqref
        ]


def main():
    """Main execution function - orchestrates workbook creation."""
    logger.info("=" * 80)
    logger.info(f"{DOCUMENT_ID} - Learning & Continuous Improvement Assessment Generator")
    logger.info(f"{CONTROL_REF}")
    logger.info("=" * 80)

    wb = create_workbook()
    styles = _STYLES

    logger.info("\n[1/10] Creating Instructions & Legend sheet...")
    create_instructions_sheet(wb["Instructions & Legend"])

    logger.info("[2/10] Creating PIR Process sheet...")
    create_pir_process_sheet(wb["PIR Process"], styles)

    logger.info("[3/10] Creating Root Cause Analysis sheet...")
    create_rca_sheet(wb["Root Cause Analysis"], styles)

    logger.info("[4/10] Creating Lessons Learned sheet...")
    create_lessons_learned_sheet(wb["Lessons Learned"], styles)

    logger.info("[5/10] Creating Control Improvements sheet...")
    create_control_improvements_sheet(wb["Control Improvements"], styles)

    logger.info("[6/10] Creating Trend Analysis sheet...")
    create_trend_analysis_sheet(wb["Trend Analysis"], styles)

    logger.info("[7/10] Creating Gap Analysis sheet...")
    create_gap_analysis_sheet(wb["Gap Analysis"], styles)

    logger.info("[8/10] Creating Evidence Register sheet...")
    create_evidence_register(wb["Evidence Register"])

    logger.info("[9/10] Creating Summary Dashboard sheet...")
    create_summary_dashboard_sheet(wb["Summary Dashboard"], styles)

    logger.info("[10/10] Creating Approval Sign-Off sheet...")
    create_approval_sheet(wb["Approval Sign-Off"])

    finalize_validations(wb)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    output_path = _wkbk_dir / OUTPUT_FILENAME
    wb.save(output_path)
    logger.info(f"\nSUCCESS: {_wkbk_dir / OUTPUT_FILENAME}")
    logger.info("\nWorkbook Structure:")
    logger.info("  - 10 sheets (Instructions through Approval)")
    logger.info("  - PIR inventory (50 rows)")
    logger.info("  - RCA inventory (30 rows for Critical/High)")
    logger.info("  - Lessons Learned log (50 rows)")
    logger.info("  - Remediation actions (70 rows)")
    logger.info("  - KPI tracking (8 pre-defined KPIs)")
    logger.info("  - Gap analysis (30 rows)")
    logger.info("  - Evidence register (50 rows)")
    logger.info("\nNext steps:")
    logger.info("  1) Complete document information in Instructions & Legend")
    logger.info("  2) Review PIRs from last 12 months in PIR Process sheet")
    logger.info("  3) Assess RCA quality for Critical/High incidents")
    logger.info("  4) Verify Lessons Learned documentation and distribution")
    logger.info("  5) Track remediation actions in Control Improvements")
    logger.info("  6) Complete Trend Analysis with current KPI values")
    logger.info("  7) Review Gap Analysis and prioritize remediation")
    logger.info("  8) Link evidence in Evidence Register")
    logger.info("  9) Review Summary Dashboard")
    logger.info("  10) Complete Approval Sign-Off workflow")
    logger.info("\nEstimated completion time: 6-10 hours")
    logger.info("\n" + "=" * 80)


if __name__ == "__main__":
    sys.exit(main())
# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
