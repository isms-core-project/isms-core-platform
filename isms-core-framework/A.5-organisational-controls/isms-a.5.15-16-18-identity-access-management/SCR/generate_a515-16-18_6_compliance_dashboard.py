#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.15-16-18.S6 - IAM Compliance Dashboard Generator
================================================================================

ISO/IEC 27001:2022 Controls A.5.15, A.5.16, A.5.18: Identity and Access Management
Assessment Workbook 6 of 6: IAM Governance Compliance Dashboard

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANIZATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organization's specific IAM governance structure, compliance targets,
and executive reporting requirements.

Key customization areas:
1. Compliance score weights (adjust domain weighting for your risk profile)
2. KPI targets (align with your organizational policy and risk tolerance)
3. Maturity level thresholds (customize for your certification timeline)
4. Evidence requirements (specific to your audit framework)
5. Trend analysis periods (monthly, quarterly, annual as appropriate)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.15-16-18 IAM Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel dashboard workbook that consolidates
IAM compliance metrics from all five assessment domains (S1-S5) into a single
executive view.

**Purpose:**
Enables executive oversight of IAM compliance posture, supporting governance
reporting, audit readiness assessment, and continuous improvement tracking
across all ISO 27001:2022 Controls A.5.15, A.5.16, and A.5.18 requirements.

**Assessment Scope:**
- Composite IAM compliance score (weighted average across S1-S5)
- Individual domain compliance scores with RAG status
- Consolidated gap register from all assessments
- Key performance indicators (KPIs) with targets and actuals
- Evidence summary for audit readiness
- Trend analysis (if historical data available)
- Certification readiness assessment

**Generated Workbook Structure:**
1. Instructions & Legend - Dashboard overview and status legends
2. Executive_Summary - One-page IAM compliance overview
3. Domain_Compliance - Individual S1-S5 domain scores
4. Gap_Analysis - Consolidated gap register
5. KPI_Dashboard - Key performance indicators
6. Evidence_Summary - Audit evidence catalogue
7. Trend_Analysis - Historical compliance tracking
8. Certification_Readiness - ISO 27001 audit preparation
9. Approval_Sign_Off - Three-level approval workflow

**Key Features:**
- Data validation with status dropdown lists
- Conditional formatting for RAG (Red/Amber/Green) status
- Composite score calculation with configurable weighting
- Gap prioritization by risk level
- Evidence completeness tracking
- Multi-stakeholder approval workflow

**Integration:**
This dashboard consolidates data from assessments S1-S5:
- S1: User Inventory & Lifecycle Compliance
- S2: Access Rights Matrix
- S3: Access Review Results
- S4: Role Definition & SoD Compliance
- S5: Lifecycle Compliance Detailed

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl

    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - random (standard library - for sample data generation)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a515-16-18_6_compliance_dashboard.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a515-16-18_6_compliance_dashboard.py --output /path/to/dir

    # Generate with specific date suffix
    python3 generate_a515-16-18_6_compliance_dashboard.py --date 20260203

Output:
    File: ISMS-IMP-A.5.15-16-18.S6_IAM_Compliance_Dashboard_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Ensure all S1-S5 assessments are completed and current
    2. Extract metrics from each source assessment
    3. Populate Executive Summary with composite score
    4. Enter individual domain scores from S1-S5
    5. Consolidate gaps from all assessments
    6. Update KPIs with current actuals
    7. Add evidence from all assessments
    8. Calculate trends (if historical data available)
    9. Assess certification readiness
    10. Obtain stakeholder approvals
    11. Distribute to executive leadership

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.5.15, A.5.16, A.5.18
Assessment Domain:    6 of 6 (IAM Governance Compliance Dashboard)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organization] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.15-16-18: Identity Management Policy (All Sections)
    - ISMS-IMP-A.5.15-16-18.S1: User Inventory Assessment
    - ISMS-IMP-A.5.15-16-18.S2: Access Rights Matrix Assessment
    - ISMS-IMP-A.5.15-16-18.S3: Access Review Results Assessment
    - ISMS-IMP-A.5.15-16-18.S4: Role Definition & SoD Compliance Assessment
    - ISMS-IMP-A.5.15-16-18.S5: Lifecycle Compliance Detailed Assessment

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements consolidated dashboard framework per ISMS-IMP-A.5.15-16-18.S6
    - Supports executive governance reporting and audit readiness

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Dashboard is Only as Good as Source Data:**
This dashboard consolidates metrics from S1-S5 assessments. If source
assessments are stale, incomplete, or inaccurate, the dashboard will
reflect those deficiencies. Always verify source assessments are current
(within 30 days) before generating the dashboard.

**Composite Score Weighting:**
The default composite score uses equal weighting (20% each domain).
Organizations may adjust weighting based on risk profile:
- Higher weight for access rights (S2) and lifecycle (S5) = security focus
- Higher weight for access review (S3) = audit focus
- Higher weight for RBAC/SoD (S4) = compliance focus

**Executive Communication:**
The Executive Summary (Sheet 2) should be understandable by non-technical
executives within 5 minutes. Use RAG status consistently. Reserve technical
detail for supporting sheets.

**Certification Readiness:**
Be conservative in readiness assessment. If ANY critical gaps remain open,
status should be "Not Ready". Auditors will find gaps; better to find them first.

**Data Protection:**
Dashboard contains aggregated compliance data which may be sensitive:
- Compliance posture information
- Gap and vulnerability information
- Organizational risk exposure

Handle in accordance with information classification requirements.

**Maintenance:**
Review and update dashboard:
- Monthly: Refresh metrics from S1-S5 assessments
- Quarterly: Executive report preparation
- Annually: Comprehensive review with trend analysis
- Ad-hoc: When significant changes occur in any domain

================================================================================
"""

import sys
from datetime import datetime, timedelta
from pathlib import Path
import random

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as e:
    logger.error(f"ERROR: Required library not found: {e}")
    logger.info("Install required libraries: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.15-16-18.S6"
WORKBOOK_NAME = "IAM Compliance Dashboard"
CONTROL_ID = "A.5.15, A.5.16, A.5.18"
CONTROL_NAME = "Identity and Access Management"
CONTROL_REF = f"ISO/IEC 27001:2022 - Controls {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")      # Swiss format
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")   # For filenames

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"


# =============================================================================
# CONFIGURATION
# =============================================================================

# Sample data counts
GAP_COUNT = 25
KPI_COUNT = 12
EVIDENCE_COUNT = 30
TREND_PERIODS = 6

# Domain score targets (CUSTOMIZE: adjust to your organisation's targets)
TARGET_SCORE = 85  # Default target percentage

# Composite score weights (CUSTOMIZE: adjust based on risk profile)
# Equal weighting - 20% each domain
DOMAIN_WEIGHTS = {
    "S1": 0.20,  # User Inventory & Lifecycle
    "S2": 0.20,  # Access Rights Matrix
    "S3": 0.20,  # Access Review Results
    "S4": 0.20,  # Role Definition & SoD
    "S5": 0.20,  # Lifecycle Compliance
}


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    medium = Side(style="medium")

    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="002060", end_color="002060", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data_center": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "compliant": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "non_compliant": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "warning": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "score_large": {
            "font": Font(name="Calibri", size=24, bold=True),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }

    return styles


def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = style_dict["font"]
    if "fill" in style_dict:
        cell.fill = style_dict["fill"]
    if "alignment" in style_dict:
        cell.alignment = style_dict["alignment"]
    if "border" in style_dict:
        cell.border = style_dict["border"]


# =============================================================================
# SAMPLE DATA GENERATORS
# =============================================================================

def generate_domain_scores():
    """Generate sample domain scores for S1-S5."""
    domains = [
        {
            "id": "S1",
            "name": "User Inventory & Lifecycle",
            "score": random.randint(78, 92),
            "key_metrics": "Users: 1'250 | Orphaned: 15 (1.2%) | Provisioning SLA: 92%",
            "findings": "3 late provisioning events, 15 orphaned accounts detected"
        },
        {
            "id": "S2",
            "name": "Access Rights Matrix",
            "score": random.randint(72, 88),
            "key_metrics": "Access grants: 3'500 | Documented: 85% | Privileged: 45",
            "findings": "15% access undocumented, 8 users with excessive access"
        },
        {
            "id": "S3",
            "name": "Access Review Results",
            "score": random.randint(85, 98),
            "key_metrics": "Reviews: 450 | Completed: 95% | Revoked: 12%",
            "findings": "High completion rate, 54 access rights revoked"
        },
        {
            "id": "S4",
            "name": "Role Definition & SoD",
            "score": random.randint(68, 82),
            "key_metrics": "Roles: 35 | RBAC: 72% | SoD Violations: 5",
            "findings": "5 SoD violations pending remediation, RBAC adoption below target"
        },
        {
            "id": "S5",
            "name": "Lifecycle Compliance",
            "score": random.randint(75, 90),
            "key_metrics": "Joiners: 85% on-time | Leavers: 92% on-time | Movers: 78%",
            "findings": "Mover access updates delayed, 3 expired contractors"
        },
    ]
    return domains


def generate_gaps(count):
    """Generate sample gap data from all domains."""
    categories = [
        ("Provisioning", "S1"),
        ("Deprovisioning", "S1"),
        ("Orphaned Accounts", "S1"),
        ("Access Documentation", "S2"),
        ("Privileged Access", "S2"),
        ("Excessive Access", "S2"),
        ("Access Review", "S3"),
        ("Attestation", "S3"),
        ("RBAC Adoption", "S4"),
        ("SoD Violation", "S4"),
        ("Joiner SLA", "S5"),
        ("Leaver SLA", "S5"),
        ("Contractor Expiry", "S5"),
    ]

    risk_levels = ["Critical", "High", "Medium", "Low"]
    statuses = ["Open", "In Progress", "Closed"]
    owners = ["IAM Manager", "Security Team", "IT Operations", "HR Manager", "CISO"]

    gaps = []
    for i in range(count):
        cat, source = random.choice(categories)
        risk = random.choices(risk_levels, weights=[10, 25, 40, 25])[0]
        status = random.choices(statuses, weights=[40, 35, 25])[0]

        gaps.append({
            "gap_id": f"GAP-{i+1:03d}",
            "source": source,
            "category": cat,
            "description": f"{cat} gap requiring attention",
            "risk_level": risk,
            "affected": random.randint(1, 25),
            "root_cause": f"Process gap in {cat.lower()} procedures",
            "remediation": f"Implement automated {cat.lower()} controls",
            "owner": random.choice(owners),
            "due_date": (datetime.now() + timedelta(days=random.randint(7, 90))).strftime("%d.%m.%Y"),
            "status": status,
        })

    # Sort by risk level (Critical first)
    risk_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    gaps.sort(key=lambda x: risk_order[x["risk_level"]])

    return gaps


def generate_kpis():
    """Generate sample KPI data."""
    kpis = [
        ("KPI-001", "User Inventory Completeness", "100%", f"{random.randint(95, 100)}%", "S1"),
        ("KPI-002", "Provisioning On-Time Rate", ">= 95%", f"{random.randint(85, 98)}%", "S1, S5"),
        ("KPI-003", "Deprovisioning On-Time Rate", ">= 98%", f"{random.randint(88, 99)}%", "S1, S5"),
        ("KPI-004", "Orphaned Account Percentage", "<= 1%", f"{random.uniform(0.5, 2.5):.1f}%", "S1"),
        ("KPI-005", "Access Documentation Completeness", ">= 90%", f"{random.randint(78, 95)}%", "S2"),
        ("KPI-006", "Privileged Access Reviewed", "100%", f"{random.randint(85, 100)}%", "S2, S3"),
        ("KPI-007", "Access Review Completion Rate", ">= 95%", f"{random.randint(88, 99)}%", "S3"),
        ("KPI-008", "RBAC Adoption Rate", ">= 80%", f"{random.randint(65, 88)}%", "S4"),
        ("KPI-009", "Open SoD Violations", "0", f"{random.randint(0, 8)}", "S4"),
        ("KPI-010", "Contractor Expiration Compliance", "100%", f"{random.randint(85, 100)}%", "S5"),
        ("KPI-011", "Gap Closure Rate", ">= 80%", f"{random.randint(60, 90)}%", "All"),
        ("KPI-012", "Evidence Completeness", ">= 95%", f"{random.randint(80, 98)}%", "All"),
    ]
    return kpis


def generate_evidence(count):
    """Generate sample evidence data."""
    evidence_types = [
        ("User Inventory", "Spreadsheet", "S1"),
        ("Lifecycle Metrics", "Process Records", "S1"),
        ("Orphaned Scan", "Scan Results", "S1"),
        ("Access Matrix", "Spreadsheet", "S2"),
        ("Privileged Access List", "Access Report", "S2"),
        ("Group Memberships", "Export Data", "S2"),
        ("Review Results", "Assessment Report", "S3"),
        ("Attestation Records", "Approval Records", "S3"),
        ("Role Catalogue", "Documentation", "S4"),
        ("SoD Matrix", "Spreadsheet", "S4"),
        ("JML Event Log", "System Logs", "S5"),
        ("HR Integration", "Configuration", "S5"),
        ("IAM Policy", "Policy Document", "All"),
        ("Procedure Guide", "Process Document", "All"),
    ]

    completeness = ["Complete", "Partial", "Missing"]
    reviewers = ["IAM Manager", "Security Manager", "IT Manager", "CISO", "Audit Team"]

    evidence = []
    for i in range(count):
        ev_name, ev_type, source = random.choice(evidence_types)
        compl = random.choices(completeness, weights=[70, 25, 5])[0]

        iso_req = random.choice(["A.5.15", "A.5.16", "A.5.18"])

        evidence.append({
            "ev_id": f"EV-{i+1:03d}",
            "source": source,
            "iso_req": iso_req,
            "req_area": ev_name,
            "ev_type": ev_type,
            "location": f"{source} Assessment Workbook" if source != "All" else "ISMS Evidence Library",
            "date": GENERATED_DATE,
            "completeness": compl,
            "reviewer": random.choice(reviewers),
        })

    return evidence


def generate_trend_data(periods):
    """Generate sample historical trend data."""
    trends = []
    base_date = datetime.now()

    # Start with lower scores and improve over time
    base_scores = {
        "S1": random.randint(70, 78),
        "S2": random.randint(65, 75),
        "S3": random.randint(75, 82),
        "S4": random.randint(60, 70),
        "S5": random.randint(68, 76),
    }

    for i in range(periods):
        month_date = base_date - timedelta(days=30 * (periods - 1 - i))
        period_name = month_date.strftime("%b %Y")

        # Gradually improve scores
        improvement = i * random.uniform(1.5, 3.0)

        scores = {
            "S1": min(100, base_scores["S1"] + improvement + random.uniform(-2, 2)),
            "S2": min(100, base_scores["S2"] + improvement + random.uniform(-2, 2)),
            "S3": min(100, base_scores["S3"] + improvement + random.uniform(-2, 2)),
            "S4": min(100, base_scores["S4"] + improvement + random.uniform(-2, 2)),
            "S5": min(100, base_scores["S5"] + improvement + random.uniform(-2, 2)),
        }

        composite = sum(scores[d] * DOMAIN_WEIGHTS[d] for d in scores)

        trends.append({
            "period": period_name,
            "composite": round(composite, 1),
            "S1": round(scores["S1"], 1),
            "S2": round(scores["S2"], 1),
            "S3": round(scores["S3"], 1),
            "S4": round(scores["S4"], 1),
            "S5": round(scores["S5"], 1),
        })

    return trends


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================

def create_instructions_sheet(wb, styles):
    """Create Sheet 1: Instructions & Legend."""
    ws = wb.create_sheet("Instructions & Legend", 0)

    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = (
        f"{DOCUMENT_ID} - {WORKBOOK_NAME}\n"
        f"{CONTROL_REF}"
    )
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 45

    # Document metadata
    ws["A3"] = "Document ID:"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = DOCUMENT_ID

    ws["A4"] = "Assessment Area:"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = "IAM Governance Compliance Dashboard"

    ws["A5"] = "Version:"
    ws["A5"].font = Font(bold=True)
    ws["B5"] = "1.0"

    ws["A6"] = "Generated:"
    ws["A6"].font = Font(bold=True)
    ws["B6"] = datetime.now().strftime("%d.%m.%Y %H:%M")

    # Purpose section
    row = 8
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Purpose"
    apply_style(ws[f"A{row}"], styles["subheader"])

    purpose_text = """
This dashboard consolidates IAM compliance metrics from all five assessment domains (S1-S5)
into a single executive view. It provides:
- Composite IAM compliance score across all domains
- Individual domain compliance scores with RAG status
- Consolidated gap register with risk prioritisation
- Key performance indicators (KPIs) tracking
- Evidence summary for audit readiness
- Trend analysis for continuous improvement
- Certification readiness assessment
    """.strip()

    row += 1
    ws.merge_cells(f"A{row}:H{row+7}")
    ws[f"A{row}"] = purpose_text
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 130

    # Source assessments
    row += 9
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Source Assessments"
    apply_style(ws[f"A{row}"], styles["subheader"])

    row += 1
    source_headers = ["Assessment", "Document ID", "Description"]
    for col, header in enumerate(source_headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])

    sources = [
        ("S1", "ISMS-IMP-A.5.15-16-18.S1", "User Inventory & Lifecycle Compliance"),
        ("S2", "ISMS-IMP-A.5.15-16-18.S2", "Access Rights Matrix"),
        ("S3", "ISMS-IMP-A.5.15-16-18.S3", "Access Review Results"),
        ("S4", "ISMS-IMP-A.5.15-16-18.S4", "Role Definition & SoD Compliance"),
        ("S5", "ISMS-IMP-A.5.15-16-18.S5", "Lifecycle Compliance Detailed"),
    ]

    for assess, doc_id, desc in sources:
        row += 1
        ws.cell(row=row, column=1).value = assess
        ws.cell(row=row, column=2).value = doc_id
        ws.cell(row=row, column=3).value = desc
        for col in range(1, 4):
            apply_style(ws.cell(row=row, column=col), styles["data"])

    # Status legend
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Status Legend"
    apply_style(ws[f"A{row}"], styles["subheader"])

    row += 1
    ws[f"A{row}"] = "Compliant"
    apply_style(ws[f"A{row}"], styles["compliant"])
    ws[f"B{row}"] = "Meets or exceeds target (>= 85%)"

    row += 1
    ws[f"A{row}"] = "Warning"
    apply_style(ws[f"A{row}"], styles["warning"])
    ws[f"B{row}"] = "Below target but acceptable (60-84%)"

    row += 1
    ws[f"A{row}"] = "Non-Compliant"
    apply_style(ws[f"A{row}"], styles["non_compliant"])
    ws[f"B{row}"] = "Significantly below target (< 60%)"

    # Maturity levels
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Maturity Levels"
    apply_style(ws[f"A{row}"], styles["subheader"])

    row += 1
    maturity_headers = ["Level", "Score Range", "Description"]
    for col, header in enumerate(maturity_headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])

    maturity_levels = [
        ("Optimised", "90-100%", "Best-in-class, continuous improvement"),
        ("Managed", "75-89%", "Strong controls, minor gaps"),
        ("Defined", "60-74%", "Documented processes, significant gaps"),
        ("Developing", "40-59%", "Basic controls, major gaps"),
        ("Initial", "< 40%", "Ad-hoc processes, critical gaps"),
    ]

    for level, score_range, desc in maturity_levels:
        row += 1
        ws.cell(row=row, column=1).value = level
        ws.cell(row=row, column=2).value = score_range
        ws.cell(row=row, column=3).value = desc
        for col in range(1, 4):
            apply_style(ws.cell(row=row, column=col), styles["data"])

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 50

    return ws


def create_executive_summary_sheet(wb, styles, domains):
    """Create Sheet 2: Executive Summary."""
    ws = wb.create_sheet("Executive_Summary")

    # Title
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "Executive Summary - IAM Compliance Dashboard"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Dashboard Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Assessment Period: {datetime.now().strftime('%B %Y')}"
    ws["A3"].font = Font(italic=True)

    # Calculate composite score
    composite_score = sum(d["score"] * DOMAIN_WEIGHTS[d["id"]] for d in domains)

    # Determine maturity level
    if composite_score >= 90:
        maturity = "Optimised"
    elif composite_score >= 75:
        maturity = "Managed"
    elif composite_score >= 60:
        maturity = "Defined"
    elif composite_score >= 40:
        maturity = "Developing"
    else:
        maturity = "Initial"

    # Composite score section
    row = 5
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "Overall IAM Compliance Score"
    apply_style(ws[f"A{row}"], styles["subheader"])

    row += 1
    ws.merge_cells(f"A{row}:B{row+1}")
    score_cell = ws[f"A{row}"]
    score_cell.value = f"{composite_score:.0f}%"
    apply_style(score_cell, styles["score_large"])

    # Apply colour based on score
    if composite_score >= 85:
        score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif composite_score >= 60:
        score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.row_dimensions[row].height = 35
    ws.row_dimensions[row + 1].height = 35

    ws[f"C{row}"] = "Maturity Level:"
    ws[f"C{row}"].font = Font(bold=True)
    ws[f"D{row}"] = maturity
    ws[f"D{row}"].font = Font(bold=True, size=14)

    ws[f"C{row+1}"] = "Target Score:"
    ws[f"C{row+1}"].font = Font(bold=True)
    ws[f"D{row+1}"] = f"{TARGET_SCORE}%"

    # Domain scores summary
    row += 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "Domain Compliance Scores"
    apply_style(ws[f"A{row}"], styles["subheader"])

    row += 1
    headers = ["Domain", "Score", "Target", "Gap", "Status"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])

    for domain in domains:
        row += 1
        gap = domain["score"] - TARGET_SCORE
        if domain["score"] >= 85:
            status = "Compliant"
        elif domain["score"] >= 60:
            status = "Warning"
        else:
            status = "Non-Compliant"

        ws.cell(row=row, column=1).value = f"{domain['id']}: {domain['name']}"
        ws.cell(row=row, column=2).value = f"{domain['score']}%"
        ws.cell(row=row, column=3).value = f"{TARGET_SCORE}%"
        ws.cell(row=row, column=4).value = f"{gap:+d}%"
        ws.cell(row=row, column=5).value = status

        status_cell = ws.cell(row=row, column=5)
        if status == "Compliant":
            apply_style(status_cell, styles["compliant"])
        elif status == "Warning":
            apply_style(status_cell, styles["warning"])
        else:
            apply_style(status_cell, styles["non_compliant"])

        for col in range(1, 5):
            apply_style(ws.cell(row=row, column=col), styles["data"])

    # Key findings section
    row += 3
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "Key Findings"
    apply_style(ws[f"A{row}"], styles["subheader"])

    # Strengths
    row += 1
    ws[f"A{row}"] = "Strengths:"
    ws[f"A{row}"].font = Font(bold=True)
    sorted_domains = sorted(domains, key=lambda x: x["score"], reverse=True)
    ws[f"B{row}"] = f"{sorted_domains[0]['name']} ({sorted_domains[0]['score']}%)"

    row += 1
    ws[f"B{row}"] = f"{sorted_domains[1]['name']} ({sorted_domains[1]['score']}%)"

    # Weaknesses
    row += 1
    ws[f"A{row}"] = "Weaknesses:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = f"{sorted_domains[-1]['name']} ({sorted_domains[-1]['score']}%)"

    row += 1
    ws[f"B{row}"] = f"{sorted_domains[-2]['name']} ({sorted_domains[-2]['score']}%)"

    # Recommendations
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "Recommendations"
    apply_style(ws[f"A{row}"], styles["subheader"])

    recommendations = [
        ("Priority 1:", "Address SoD violations in S4 domain (5 violations pending)"),
        ("Priority 2:", "Improve access documentation completeness in S2 (target 90%)"),
        ("Priority 3:", "Accelerate RBAC adoption to meet 80% target"),
        ("Strategic:", "Implement automated HR-to-identity integration for lifecycle events"),
    ]

    for priority, recommendation in recommendations:
        row += 1
        ws[f"A{row}"] = priority
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = recommendation

    # Certification readiness
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "Certification Readiness"
    apply_style(ws[f"A{row}"], styles["subheader"])

    row += 1
    ws[f"A{row}"] = "Status:"
    ws[f"A{row}"].font = Font(bold=True)

    if composite_score >= 85:
        readiness = "Ready"
        readiness_style = styles["compliant"]
    elif composite_score >= 70:
        readiness = "Mostly Ready"
        readiness_style = styles["warning"]
    else:
        readiness = "Not Ready"
        readiness_style = styles["non_compliant"]

    ws[f"B{row}"] = readiness
    apply_style(ws[f"B{row}"], readiness_style)

    row += 1
    ws[f"A{row}"] = "Audit Blockers:"
    ws[f"A{row}"].font = Font(bold=True)
    blockers = sum(1 for d in domains if d["score"] < 60)
    ws[f"B{row}"] = f"{blockers} critical domain(s) below threshold"

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15

    return ws


def create_domain_compliance_sheet(wb, styles, domains):
    """Create Sheet 3: Domain Compliance Scores."""
    ws = wb.create_sheet("Domain_Compliance")

    # Title
    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = "Domain Compliance Scores - S1 through S5"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Assessment Period: {datetime.now().strftime('%B %Y')}"
    ws["A2"].font = Font(italic=True)

    # Column headers
    row = 4
    headers = ["Domain ID", "Domain Name", "Score", "Target", "Gap", "Status", "Trend", "Key Metrics", "Findings Summary"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])

    trends = ["Improving", "Stable", "Declining"]

    for domain in domains:
        row += 1
        gap = domain["score"] - TARGET_SCORE
        if domain["score"] >= 85:
            status = "Compliant"
        elif domain["score"] >= 60:
            status = "Warning"
        else:
            status = "Non-Compliant"

        trend = random.choices(trends, weights=[50, 35, 15])[0]

        ws.cell(row=row, column=1).value = domain["id"]
        ws.cell(row=row, column=2).value = domain["name"]
        ws.cell(row=row, column=3).value = f"{domain['score']}%"
        ws.cell(row=row, column=4).value = f"{TARGET_SCORE}%"
        ws.cell(row=row, column=5).value = f"{gap:+d}%"
        ws.cell(row=row, column=6).value = status
        ws.cell(row=row, column=7).value = trend
        ws.cell(row=row, column=8).value = domain["key_metrics"]
        ws.cell(row=row, column=9).value = domain["findings"]

        # Status formatting
        status_cell = ws.cell(row=row, column=6)
        if status == "Compliant":
            apply_style(status_cell, styles["compliant"])
        elif status == "Warning":
            apply_style(status_cell, styles["warning"])
        else:
            apply_style(status_cell, styles["non_compliant"])

        # Trend formatting
        trend_cell = ws.cell(row=row, column=7)
        if trend == "Improving":
            apply_style(trend_cell, styles["compliant"])
        elif trend == "Stable":
            apply_style(trend_cell, styles["warning"])
        else:
            apply_style(trend_cell, styles["non_compliant"])

        for col in range(1, 10):
            if col not in [6, 7]:
                apply_style(ws.cell(row=row, column=col), styles["data"])

    # Column widths
    widths = {"A": 12, "B": 30, "C": 10, "D": 10, "E": 10, "F": 15, "G": 12, "H": 50, "I": 50}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"

    return ws


def create_gap_analysis_sheet(wb, styles, gaps):
    """Create Sheet 4: Consolidated Gap Analysis."""
    ws = wb.create_sheet("Gap_Analysis")

    # Title
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = "Consolidated Gap Analysis - All IAM Domains"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)

    # Gap summary
    critical_count = sum(1 for g in gaps if g["risk_level"] == "Critical")
    high_count = sum(1 for g in gaps if g["risk_level"] == "High")
    open_count = sum(1 for g in gaps if g["status"] == "Open")

    ws["A3"] = f"Total Gaps: {len(gaps)} | Critical: {critical_count} | High: {high_count} | Open: {open_count}"
    ws["A3"].font = Font(bold=True)

    # Column headers
    row = 5
    headers = ["Gap ID", "Source", "Category", "Description", "Risk Level", "Affected",
               "Root Cause", "Remediation Plan", "Owner", "Due Date", "Status", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])

    for gap in gaps:
        row += 1
        ws.cell(row=row, column=1).value = gap["gap_id"]
        ws.cell(row=row, column=2).value = gap["source"]
        ws.cell(row=row, column=3).value = gap["category"]
        ws.cell(row=row, column=4).value = gap["description"]
        ws.cell(row=row, column=5).value = gap["risk_level"]
        ws.cell(row=row, column=6).value = gap["affected"]
        ws.cell(row=row, column=7).value = gap["root_cause"]
        ws.cell(row=row, column=8).value = gap["remediation"]
        ws.cell(row=row, column=9).value = gap["owner"]
        ws.cell(row=row, column=10).value = gap["due_date"]
        ws.cell(row=row, column=11).value = gap["status"]

        # Risk level formatting
        risk_cell = ws.cell(row=row, column=5)
        if gap["risk_level"] == "Critical":
            risk_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            risk_cell.font = Font(bold=True, color="FFFFFF")
        elif gap["risk_level"] == "High":
            apply_style(risk_cell, styles["non_compliant"])
        elif gap["risk_level"] == "Medium":
            apply_style(risk_cell, styles["warning"])
        else:
            apply_style(risk_cell, styles["compliant"])

        # Status formatting
        status_cell = ws.cell(row=row, column=11)
        if gap["status"] == "Closed":
            apply_style(status_cell, styles["compliant"])
        elif gap["status"] == "In Progress":
            apply_style(status_cell, styles["warning"])
        else:
            apply_style(status_cell, styles["non_compliant"])

        for col in range(1, 13):
            if col not in [5, 11]:
                apply_style(ws.cell(row=row, column=col), styles["data"])

    # Column widths
    widths = {"A": 10, "B": 8, "C": 18, "D": 35, "E": 12, "F": 10,
              "G": 30, "H": 35, "I": 15, "J": 12, "K": 12, "L": 25}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A6"

    return ws


def create_kpi_dashboard_sheet(wb, styles, kpis):
    """Create Sheet 5: KPI Dashboard."""
    ws = wb.create_sheet("KPI_Dashboard")

    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "IAM Key Performance Indicators (KPIs)"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Assessment Period: {datetime.now().strftime('%B %Y')}"
    ws["A2"].font = Font(italic=True)

    # Column headers
    row = 4
    headers = ["KPI ID", "KPI Name", "Target", "Actual", "Gap", "Status", "Trend", "Source"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])

    trends = ["Improving", "Stable", "Declining"]

    for kpi_id, kpi_name, target, actual, source in kpis:
        row += 1

        # Parse target and actual for comparison
        try:
            if ">=" in target:
                target_val = float(target.replace(">=", "").replace("%", "").strip())
                actual_val = float(actual.replace("%", "").strip())
                gap = actual_val - target_val
                meets_target = actual_val >= target_val
            elif "<=" in target:
                target_val = float(target.replace("<=", "").replace("%", "").strip())
                actual_val = float(actual.replace("%", "").strip())
                gap = target_val - actual_val
                meets_target = actual_val <= target_val
            elif target == "0":
                actual_val = float(actual)
                gap = -actual_val
                meets_target = actual_val == 0
            else:
                target_val = float(target.replace("%", "").strip())
                actual_val = float(actual.replace("%", "").strip())
                gap = actual_val - target_val
                meets_target = actual_val >= target_val
        except (ValueError, AttributeError) as e:
            gap = 0
            meets_target = True

        if meets_target:
            status = "Compliant"
        elif abs(gap) <= 10:
            status = "Warning"
        else:
            status = "Non-Compliant"

        trend = random.choices(trends, weights=[50, 35, 15])[0]

        ws.cell(row=row, column=1).value = kpi_id
        ws.cell(row=row, column=2).value = kpi_name
        ws.cell(row=row, column=3).value = target
        ws.cell(row=row, column=4).value = actual
        ws.cell(row=row, column=5).value = f"{gap:+.1f}%" if "%" in str(actual) else f"{gap:+.0f}"
        ws.cell(row=row, column=6).value = status
        ws.cell(row=row, column=7).value = trend
        ws.cell(row=row, column=8).value = source

        # Status formatting
        status_cell = ws.cell(row=row, column=6)
        if status == "Compliant":
            apply_style(status_cell, styles["compliant"])
        elif status == "Warning":
            apply_style(status_cell, styles["warning"])
        else:
            apply_style(status_cell, styles["non_compliant"])

        # Trend formatting
        trend_cell = ws.cell(row=row, column=7)
        if trend == "Improving":
            apply_style(trend_cell, styles["compliant"])
        elif trend == "Stable":
            apply_style(trend_cell, styles["warning"])
        else:
            apply_style(trend_cell, styles["non_compliant"])

        for col in range(1, 9):
            if col not in [6, 7]:
                apply_style(ws.cell(row=row, column=col), styles["data"])

    # Column widths
    widths = {"A": 10, "B": 35, "C": 12, "D": 12, "E": 10, "F": 15, "G": 12, "H": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"

    return ws


def create_evidence_summary_sheet(wb, styles, evidence):
    """Create Sheet 6: Evidence Summary."""
    ws = wb.create_sheet("Evidence_Summary")

    # Title
    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = "Evidence Summary - ISO 27001 A.5.15/A.5.16/A.5.18"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Evidence Collection Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)

    # Evidence summary metrics
    complete_count = sum(1 for e in evidence if e["completeness"] == "Complete")
    partial_count = sum(1 for e in evidence if e["completeness"] == "Partial")
    missing_count = sum(1 for e in evidence if e["completeness"] == "Missing")
    completeness_pct = (complete_count / len(evidence)) * 100 if evidence else 0

    ws["A3"] = f"Total Evidence: {len(evidence)} | Complete: {complete_count} ({completeness_pct:.0f}%) | Partial: {partial_count} | Missing: {missing_count}"
    ws["A3"].font = Font(bold=True)

    # Column headers
    row = 5
    headers = ["Evidence ID", "Source", "ISO Requirement", "Requirement Area", "Evidence Type",
               "Evidence Location", "Collection Date", "Completeness", "Verified By"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])

    for ev in evidence:
        row += 1
        ws.cell(row=row, column=1).value = ev["ev_id"]
        ws.cell(row=row, column=2).value = ev["source"]
        ws.cell(row=row, column=3).value = ev["iso_req"]
        ws.cell(row=row, column=4).value = ev["req_area"]
        ws.cell(row=row, column=5).value = ev["ev_type"]
        ws.cell(row=row, column=6).value = ev["location"]
        ws.cell(row=row, column=7).value = ev["date"]
        ws.cell(row=row, column=8).value = ev["completeness"]
        ws.cell(row=row, column=9).value = ev["reviewer"]

        # Completeness formatting
        comp_cell = ws.cell(row=row, column=8)
        if ev["completeness"] == "Complete":
            apply_style(comp_cell, styles["compliant"])
        elif ev["completeness"] == "Partial":
            apply_style(comp_cell, styles["warning"])
        else:
            apply_style(comp_cell, styles["non_compliant"])

        for col in range(1, 10):
            if col != 8:
                apply_style(ws.cell(row=row, column=col), styles["data"])

    # Column widths
    widths = {"A": 12, "B": 8, "C": 15, "D": 25, "E": 18, "F": 35, "G": 15, "H": 15, "I": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A6"

    return ws


def create_trend_analysis_sheet(wb, styles, trend_data):
    """Create Sheet 7: Trend Analysis."""
    ws = wb.create_sheet("Trend_Analysis")

    # Title
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "Trend Analysis - IAM Compliance Trajectory"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Analysis Period: Last {len(trend_data)} months"
    ws["A2"].font = Font(italic=True)

    # Section header
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "Historical Compliance Scores"
    apply_style(ws[f"A{row}"], styles["subheader"])

    # Column headers
    row += 1
    headers = ["Period", "Composite", "S1", "S2", "S3", "S4", "S5"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])

    for trend in trend_data:
        row += 1
        ws.cell(row=row, column=1).value = trend["period"]
        ws.cell(row=row, column=2).value = f"{trend['composite']:.1f}%"
        ws.cell(row=row, column=3).value = f"{trend['S1']:.1f}%"
        ws.cell(row=row, column=4).value = f"{trend['S2']:.1f}%"
        ws.cell(row=row, column=5).value = f"{trend['S3']:.1f}%"
        ws.cell(row=row, column=6).value = f"{trend['S4']:.1f}%"
        ws.cell(row=row, column=7).value = f"{trend['S5']:.1f}%"

        for col in range(1, 8):
            apply_style(ws.cell(row=row, column=col), styles["data_center"])

    # Trend summary
    row += 3
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "Trend Summary"
    apply_style(ws[f"A{row}"], styles["subheader"])

    if len(trend_data) >= 2:
        current = trend_data[-1]["composite"]
        previous = trend_data[-2]["composite"]
        change = current - previous

        row += 1
        ws[f"A{row}"] = "Current Period Score:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = f"{current:.1f}%"

        row += 1
        ws[f"A{row}"] = "Previous Period Score:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = f"{previous:.1f}%"

        row += 1
        ws[f"A{row}"] = "Month-over-Month Change:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = f"{change:+.1f}%"
        if change > 0:
            apply_style(ws[f"B{row}"], styles["compliant"])
        elif change < 0:
            apply_style(ws[f"B{row}"], styles["non_compliant"])

        row += 1
        ws[f"A{row}"] = "Trend Direction:"
        ws[f"A{row}"].font = Font(bold=True)
        if change > 1:
            direction = "Improving"
            dir_style = styles["compliant"]
        elif change < -1:
            direction = "Declining"
            dir_style = styles["non_compliant"]
        else:
            direction = "Stable"
            dir_style = styles["warning"]
        ws[f"B{row}"] = direction
        apply_style(ws[f"B{row}"], dir_style)

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    for col in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 12

    ws.freeze_panes = "A6"

    return ws


def create_certification_readiness_sheet(wb, styles, domains):
    """Create Sheet 8: Certification Readiness."""
    ws = wb.create_sheet("Certification_Readiness")

    # Title
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "Certification Readiness - ISO 27001 Audit Preparation"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)

    # Column headers
    row = 4
    headers = ["Control", "Control Name", "Source Assessments", "Evidence Status", "Gap Status", "Readiness", "Audit Blockers"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])

    # Control assessments
    controls = [
        ("A.5.15", "Access Control", "S2, S3", "Complete", "Minor Gaps", "Mostly Ready", ""),
        ("A.5.16", "Identity Management", "S1, S5", "Complete", "Minor Gaps", "Mostly Ready", ""),
        ("A.5.18", "Access Rights", "S2, S4", "Partial", "Significant Gaps", "Not Ready", "SoD violations pending"),
    ]

    for control_id, control_name, sources, evidence, gaps, readiness, blockers in controls:
        row += 1
        ws.cell(row=row, column=1).value = control_id
        ws.cell(row=row, column=2).value = control_name
        ws.cell(row=row, column=3).value = sources
        ws.cell(row=row, column=4).value = evidence
        ws.cell(row=row, column=5).value = gaps
        ws.cell(row=row, column=6).value = readiness
        ws.cell(row=row, column=7).value = blockers

        # Evidence status formatting
        ev_cell = ws.cell(row=row, column=4)
        if evidence == "Complete":
            apply_style(ev_cell, styles["compliant"])
        elif evidence == "Partial":
            apply_style(ev_cell, styles["warning"])
        else:
            apply_style(ev_cell, styles["non_compliant"])

        # Readiness formatting
        ready_cell = ws.cell(row=row, column=6)
        if readiness == "Ready":
            apply_style(ready_cell, styles["compliant"])
        elif readiness == "Mostly Ready":
            apply_style(ready_cell, styles["warning"])
        else:
            apply_style(ready_cell, styles["non_compliant"])

        for col in range(1, 8):
            if col not in [4, 6]:
                apply_style(ws.cell(row=row, column=col), styles["data"])

    # Overall readiness
    row += 3
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "Overall Certification Readiness"
    apply_style(ws[f"A{row}"], styles["subheader"])

    # Calculate composite score for readiness
    composite = sum(d["score"] * DOMAIN_WEIGHTS[d["id"]] for d in domains)

    if composite >= 85:
        overall_readiness = "Ready"
        readiness_style = styles["compliant"]
    elif composite >= 70:
        overall_readiness = "Mostly Ready"
        readiness_style = styles["warning"]
    else:
        overall_readiness = "Not Ready"
        readiness_style = styles["non_compliant"]

    row += 1
    ws[f"A{row}"] = "Overall Readiness:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = overall_readiness
    apply_style(ws[f"B{row}"], readiness_style)

    row += 1
    ws[f"A{row}"] = "Total Audit Blockers:"
    ws[f"A{row}"].font = Font(bold=True)
    blocker_count = sum(1 for c in controls if c[6])
    ws[f"B{row}"] = f"{blocker_count}"

    row += 1
    ws[f"A{row}"] = "Estimated Remediation:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "4-6 weeks" if blocker_count > 0 else "None required"

    # Column widths
    widths = {"A": 12, "B": 25, "C": 15, "D": 15, "E": 18, "F": 15, "G": 35}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"

    return ws


def create_approval_sheet(wb, styles):
    """Create Sheet 9: Approval Sign-Off."""
    ws = wb.create_sheet("Approval_Sign_Off")

    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Dashboard Approval & Sign-Off"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Dashboard Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)

    # Approval section
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "3-Level Approval Process"
    apply_style(ws[f"A{row}"], styles["subheader"])

    # Column headers
    row += 1
    headers = ["Approval Level", "Role", "Name", "Signature", "Date", "Status"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])

    approvals = [
        ("Level 1: Prepared By", "IAM Manager", "[Name]", "", "", "Pending"),
        ("Level 2: Reviewed By", "Security Manager", "[Name]", "", "", "Pending"),
        ("Level 3: Approved By", "CISO", "[Name]", "", "", "Pending"),
    ]

    for level, role, name, sig, date, status in approvals:
        row += 1
        ws.cell(row=row, column=1).value = level
        ws.cell(row=row, column=2).value = role
        ws.cell(row=row, column=3).value = name
        ws.cell(row=row, column=4).value = sig
        ws.cell(row=row, column=5).value = date
        ws.cell(row=row, column=6).value = status

        for col in range(1, 7):
            apply_style(ws.cell(row=row, column=col), styles["data"])

    # Distribution section
    row += 3
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Distribution"
    apply_style(ws[f"A{row}"], styles["subheader"])

    row += 1
    ws[f"A{row}"] = "Approved for Distribution:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "No"

    row += 1
    ws[f"A{row}"] = "Distribution Date:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "[Date]"

    row += 1
    ws[f"A{row}"] = "Distribution List:"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = "CISO, Security Management, Internal Audit, Executive Leadership"

    # Column widths
    widths = {"A": 25, "B": 20, "C": 20, "D": 20, "E": 15, "F": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    return ws


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main function to generate the workbook."""
    logger.info("=" * 80)
    logger.info(f"Generating: {WORKBOOK_NAME}")
    logger.info(f"Controls: {CONTROL_REF}")
    logger.info("=" * 80)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Setup styles
    styles = setup_styles()

    # Generate sample data
    logger.info("\nGenerating sample data...")
    domains = generate_domain_scores()
    gaps = generate_gaps(GAP_COUNT)
    kpis = generate_kpis()
    evidence = generate_evidence(EVIDENCE_COUNT)
    trend_data = generate_trend_data(TREND_PERIODS)

    # Create sheets
    logger.info("\nCreating sheets...")

    logger.info("  1/9 Creating Instructions & Legend...")
    create_instructions_sheet(wb, styles)

    logger.info("  2/9 Creating Executive Summary...")
    create_executive_summary_sheet(wb, styles, domains)

    logger.info("  3/9 Creating Domain Compliance...")
    create_domain_compliance_sheet(wb, styles, domains)

    logger.info("  4/9 Creating Gap Analysis...")
    create_gap_analysis_sheet(wb, styles, gaps)

    logger.info("  5/9 Creating KPI Dashboard...")
    create_kpi_dashboard_sheet(wb, styles, kpis)

    logger.info("  6/9 Creating Evidence Summary...")
    create_evidence_summary_sheet(wb, styles, evidence)

    logger.info("  7/9 Creating Trend Analysis...")
    create_trend_analysis_sheet(wb, styles, trend_data)

    logger.info("  8/9 Creating Certification Readiness...")
    create_certification_readiness_sheet(wb, styles, domains)

    logger.info("  9/9 Creating Approval Sign-Off...")
    create_approval_sheet(wb, styles)

    # Save workbook
    output_path = Path.cwd() / OUTPUT_FILENAME

    logger.info(f"\nSaving workbook: {OUTPUT_FILENAME}")
    wb.save(output_path)

    logger.info("\n" + "=" * 80)
    logger.info("SUCCESS!")
    logger.info(f"File created: {output_path}")
    logger.info(f"Sheets: 9")
    logger.info("=" * 80)
    logger.info("\nWorkbook Contents:")
    logger.info("  - Instructions & Legend (dashboard overview)")
    logger.info("  - Executive Summary (composite score, maturity level)")
    logger.info(f"  - Domain Compliance (5 domains)")
    logger.info(f"  - Gap Analysis ({GAP_COUNT} consolidated gaps)")
    logger.info(f"  - KPI Dashboard ({len(kpis)} KPIs)")
    logger.info(f"  - Evidence Summary ({EVIDENCE_COUNT} evidence items)")
    logger.info(f"  - Trend Analysis ({TREND_PERIODS} periods)")
    logger.info("  - Certification Readiness (ISO 27001)")
    logger.info("  - Approval Sign-Off (3-level workflow)")
    logger.info("=" * 80)

    return output_path


if __name__ == "__main__":
    try:
        output_file = main()
        sys.exit(0)
    except Exception as e:
        logger.error(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# =============================================================================
# QA_VERIFIED: 2026-02-03
# QA_STATUS: PASSED - STANDARDIZATION COMPLETE (Phase 1-3)
# QA_TOOL: Claude Code Standardization
# CHANGES: Initial creation, constants, metadata headers, v1.0 versioning, logger output
# =============================================================================
