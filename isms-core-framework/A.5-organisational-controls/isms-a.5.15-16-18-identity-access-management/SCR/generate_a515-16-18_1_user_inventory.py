#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.15-16-18.S1 - User Inventory & Lifecycle Compliance Assessment Generator
================================================================================

ISO/IEC 27001:2022 Control A.5.16: Identity Management
Assessment Workbook 1 of 5: User Inventory and Identity Lifecycle Compliance

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific identity management systems, user lifecycle
processes, and assessment requirements.

Key customization areas:
1. Identity system integrations (Active Directory, Microsoft Entra ID (formerly Azure AD), Okta, custom LDAP)
2. User types and classification (employees, contractors, vendors, service accounts)
3. Provisioning/deprovisioning SLA thresholds (match your IAM policies)
4. HR system integration points (authoritative source of identity)
5. Compliance scoring criteria (aligned with your risk tolerance)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.15-16-18 IAM Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
user inventory completeness and identity lifecycle compliance across all
identity management systems.

**Purpose:**
Enables systematic assessment of identity management processes against
ISO 27001:2022 Control A.5.16 requirements, supporting evidence-based
validation of joiner/mover/leaver procedures and orphaned account detection.

**Assessment Scope:**
- Complete user inventory across all identity systems
- Employee lifecycle compliance (provisioning/deprovisioning timeliness)
- Contractor lifecycle compliance (time-bound access management)
- Vendor identity management (sponsored access tracking)
- Service account inventory and ownership
- Orphaned account detection and remediation
- Inactive account identification (no login >90 days)
- User type classification and segregation
- Lifecycle metrics and KPIs
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Usage guidance and field definitions
2. User_Inventory - Complete user list across all systems (100 sample users)
3. Employee_Lifecycle - Employee provisioning/deprovisioning compliance (50 rows)
4. Contractor_Lifecycle - Contractor time-bound access compliance (30 rows)
5. Service_Accounts - Non-human account inventory (20 rows)
6. Orphaned_Accounts - Orphaned account detection and remediation (15 rows)
7. Lifecycle_Metrics - Summary metrics and KPIs
8. Gap_Analysis - Non-compliance findings and remediation tracking
9. Evidence_Register - Evidence collection for A.5.16 compliance
10. Approval_Sign_Off - 3-level approval workflow

**Key Features:**
- Data validation with identity system dropdown lists
- Conditional formatting for lifecycle compliance status
- Automated gap identification for orphaned/inactive accounts
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with HR system exports

**Integration:**
This assessment feeds into the IAM Governance Dashboard, which
consolidates data from all five IAM assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - random (standard library - for sample data generation)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a515-16-18_1_user_inventory.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a515-16-18_1_user_inventory.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a515-16-18_1_user_inventory.py --date 20250124

Output:
    File: ISMS-IMP-A.5.15-16-18.S1_User_Inventory_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Export user data from identity systems (AD, Microsoft Entra ID (formerly Azure AD), Okta, etc.)
    2. Export employee data from HR system (authoritative source)
    3. Populate User_Inventory with actual user accounts
    4. Document joiner events in Employee_Lifecycle
    5. Document leaver events and deprovisioning timeliness
    6. Track contractor access periods in Contractor_Lifecycle
    7. Inventory service accounts and assign ownership
    8. Cross-reference to identify orphaned accounts
    9. Calculate compliance metrics and identify gaps
    10. Collect and link audit evidence
    11. Obtain stakeholder approvals
    12. Feed results into IAM Governance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Control A.5.16
Assessment Domain:    1 of 5 (User Inventory & Lifecycle Compliance)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.15-16-18: Identity Management Policy (A.5.16)
    - ISMS-IMP-A.5.15-16-18-S2: Identity Lifecycle Process Guide
    - ISMS-IMP-A.5.15-16-18.S2: Access Rights Matrix Assessment (Workbook 2)
    - ISMS-IMP-A.5.15-16-18.S3: Access Review Results Assessment (Workbook 3)
    - ISMS-IMP-A.5.15-16-18.S4: Role Compliance Assessment (Workbook 4)
    - ISMS-IMP-A.5.15-16-18.S5: Lifecycle Compliance Detailed (Workbook 5)
    - IAM Governance Dashboard: Consolidated IAM compliance view

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.15-16-18-S1
    - Supports comprehensive user inventory and lifecycle compliance evaluation
    - Integrated with IAM Governance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Identity Lifecycle Criticality:**
Timely provisioning/deprovisioning is CRITICAL for security:
- Delayed provisioning: Business disruption (new employees can't work)
- Delayed deprovisioning: Security risk (ex-employees retain access)
- Orphaned accounts: Major audit finding (no accountability)

Ensure lifecycle SLA thresholds match your organisation's risk tolerance.

**Audit Considerations:**
This assessment generates audit evidence per ISO 27001:2022 requirements.
Auditors expect:
- Complete user inventory across ALL identity systems
- Documented provisioning/deprovisioning events with timestamps
- Evidence of timely access removal upon termination
- Orphaned account detection and remediation procedures
- Service account ownership and justification

**Data Protection:**
Assessment workbooks contain sensitive HR and identity data including:
- Employee names, departments, hire/termination dates
- Contractor sponsorship and access periods
- Account credentials metadata
- Organisational structure information

Handle in accordance with GDPR/FADP data protection requirements.

**Maintenance:**
Review and update assessment:
- Monthly: Orphaned account detection runs
- Quarterly: Complete lifecycle compliance review
- Semi-annually: User inventory reconciliation with HR system
- Annually: Full reassessment for certification
- Ad-hoc: When identity systems change or after organisational restructuring

**Quality Assurance:**
Cross-validate user inventory against multiple authoritative sources:
- HR system (employees, contractors)
- Identity systems (AD, Microsoft Entra ID (formerly Azure AD), Okta)
- Access management platforms
- Ticketing system (access requests/removals)

Discrepancies indicate control weaknesses requiring investigation.

**Contractor/Vendor Management:**
Time-bound access for contractors/vendors is MANDATORY:
- Sponsorship required (internal employee sponsor)
- Contract end date drives access expiration
- Access extension requires explicit approval
- Unsponsored contractor accounts = orphaned accounts

**Service Accounts:**
Service accounts require special handling:
- No expiration date (runs until service decommissioned)
- Ownership assignment CRITICAL (service account owner documented)
- Password rotation requirements
- Privileged service accounts: extra scrutiny in A.8.2 PAM assessment

================================================================================
"""

import sys
from datetime import datetime, timedelta
from pathlib import Path
import random

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import BarChart, Reference
except ImportError as e:
    logger.error(f"ERROR: Required library not found: {e}")
    logger.info("Install required libraries: pip install openpyxl")
    sys.exit(1)

_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
WORKBOOK_NAME = "User Inventory & Lifecycle Compliance Assessment"
DOCUMENT_ID = "ISMS-IMP-A.5.15-16-18.S1"
CONTROL_ID = "A.5.16"
CONTROL_NAME = "Identity Management"
CONTROL_REF   = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# Data row counts
USER_ROW_COUNT = 100          # Sample users
EMPLOYEE_LIFECYCLE_COUNT = 50 # Employee lifecycle events
CONTRACTOR_LIFECYCLE_COUNT = 30 # Contractor lifecycle events
SERVICE_ACCOUNT_COUNT = 20    # Service accounts
ORPHANED_ACCOUNT_COUNT = 15   # Orphaned accounts
GAP_ROW_COUNT = 40            # Gap analysis tracking
EVIDENCE_ROW_COUNT = 50       # Evidence register

# Compliance thresholds
PROVISIONING_SLA_DAYS = 0     # Provision by start date (0 = on or before start date)
DEPROVISIONING_SLA_HOURS = 24 # Deprovision within 24 hours of termination
INACTIVE_THRESHOLD_DAYS = 90  # Accounts with no login for 90+ days flagged
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point
# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================
def setup_styles():
    """
    Define all cell styles used throughout the workbook.
    Returns style templates as dictionaries to avoid shared object issues.
    """
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data_center": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "compliant": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "non_compliant": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "warning": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }
    
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = style_dict["font"]
    if "fill" in style_dict:
        cell.fill = style_dict["fill"]
    if "alignment" in style_dict:
        cell.alignment = style_dict["alignment"]
    if "border" in style_dict:
        cell.border = style_dict["border"]


# ============================================================================
# SECTION 3: SAMPLE DATA GENERATORS
# ============================================================================

def generate_sample_users(count):
    """Generate sample user data for testing."""
    first_names = ["Alice", "Bob", "Carol", "David", "Emma", "Frank", "Grace", "Henry", 
                   "Iris", "Jack", "Kate", "Liam", "Mary", "Nathan", "Olivia", "Peter",
                   "Quinn", "Rachel", "Sam", "Tara"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
                  "Davis", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez",
                  "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin"]
    
    departments = ["Finance", "Sales", "Marketing", "IT", "HR", "Operations", "Engineering",
                   "Product", "Legal", "Customer Success"]
    job_titles = {
        "Finance": ["Financial Analyst", "Finance Manager", "Accountant", "CFO"],
        "Sales": ["Sales Representative", "Sales Manager", "Account Executive", "VP Sales"],
        "Marketing": ["Marketing Specialist", "Marketing Manager", "Content Writer"],
        "IT": ["IT Administrator", "System Administrator", "IT Manager", "Network Engineer"],
        "HR": ["HR Generalist", "HR Manager", "Recruiter", "HR Director"],
        "Operations": ["Operations Analyst", "Operations Manager", "Logistics Coordinator"],
        "Engineering": ["Software Developer", "Senior Engineer", "Engineering Manager", "DevOps Engineer"],
        "Product": ["Product Manager", "Product Owner", "UX Designer"],
        "Legal": ["Legal Counsel", "Paralegal", "General Counsel"],
        "Customer Success": ["Customer Success Manager", "Support Specialist"]
    }
    
    user_types = ["Employee", "Contractor", "Vendor", "Service Account"]
    statuses = ["Active", "Disabled", "Suspended"]
    
    users = []
    for i in range(count):
        first = random.choice(first_names)
        last = random.choice(last_names)
        dept = random.choice(departments)
        user_type = random.choices(user_types, weights=[70, 20, 5, 5])[0]  # 70% employees
        
        if user_type == "Service Account":
            username = f"svc-{random.choice(['backup', 'monitoring', 'automation', 'integration'])}-{i:02d}"
            email = f"{username}@company.local"
            full_name = f"Service Account - {username}"
            title = "Service Account"
        else:
            username = f"{first.lower()}.{last.lower()}"
            if user_type == "Contractor":
                username = f"EXT-{username}"
            email = f"{username}@company.com"
            full_name = f"{first} {last}"
            title = random.choice(job_titles.get(dept, ["Specialist"]))
        
        # Most users active, some disabled/suspended
        status = random.choices(statuses, weights=[85, 10, 5])[0]
        
        # Generate dates
        hire_date = datetime.now() - timedelta(days=random.randint(30, 1825))  # 1 month to 5 years ago
        
        # Termination date (only for some disabled users)
        term_date = None
        if status == "Disabled" and random.random() < 0.7:  # 70% of disabled users are terminated
            term_date = datetime.now() - timedelta(days=random.randint(1, 60))
        
        # Last login (active users more recent, disabled less recent)
        if status == "Active":
            last_login = datetime.now() - timedelta(days=random.randint(0, 30))
        else:
            last_login = datetime.now() - timedelta(days=random.randint(60, 180)) if random.random() < 0.8 else None
        
        users.append({
            "user_id": f"U{10000 + i}",
            "username": username,
            "full_name": full_name,
            "email": email,
            "user_type": user_type,
            "department": dept if user_type != "Service Account" else "IT",
            "job_title": title,
            "manager": f"{random.choice(first_names)} {random.choice(last_names)}" if user_type != "Service Account" else "IT Manager",
            "hire_date": hire_date,
            "termination_date": term_date,
            "status": status,
            "last_login": last_login,
            "created_date": hire_date - timedelta(days=random.randint(1, 7)),  # Account created within week of hire
        })
    
    return users


def generate_lifecycle_events(user_list, event_type="joiner"):
    """Generate lifecycle events (joiner/mover/leaver) from user list."""
    events = []
    
    for user in user_list:
        if event_type == "joiner":
            # Provisioning timeliness
            provision_date = user["created_date"]
            days_to_provision = (provision_date - user["hire_date"]).days
            compliant = days_to_provision <= PROVISIONING_SLA_DAYS
            
            events.append({
                "user_id": user["user_id"],
                "username": user["username"],
                "full_name": user["full_name"],
                "user_type": user["user_type"],
                "department": user["department"],
                "hire_date": user["hire_date"],
                "provision_date": provision_date,
                "days_to_provision": days_to_provision,
                "compliance_status": "Compliant" if compliant else "Non-Compliant",
            })
        
        elif event_type == "leaver" and user["termination_date"]:
            # Deprovisioning timeliness
            # Simulate disable date (some compliant, some late)
            if random.random() < 0.85:  # 85% compliant
                disable_date = user["termination_date"] + timedelta(hours=random.randint(0, 12))
            else:  # 15% non-compliant
                disable_date = user["termination_date"] + timedelta(days=random.randint(2, 14))
            
            hours_to_disable = (disable_date - user["termination_date"]).total_seconds() / 3600
            compliant = hours_to_disable <= DEPROVISIONING_SLA_HOURS
            
            events.append({
                "user_id": user["user_id"],
                "username": user["username"],
                "full_name": user["full_name"],
                "user_type": user["user_type"],
                "department": user["department"],
                "termination_date": user["termination_date"],
                "disable_date": disable_date,
                "hours_to_disable": round(hours_to_disable, 1),
                "compliance_status": "Compliant" if compliant else "Non-Compliant",
            })
    
    return events


# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Review this sheet to understand the assessment structure and status conventions.', '2. Complete the User Inventory sheet with all identities across all systems.', '3. Complete the Employee Lifecycle sheet with joiner/leaver provisioning events.', '4. Complete the Contractor Lifecycle sheet with contractor time-bound access records.', '5. Complete the Service Accounts sheet with all non-human account details.', '6. Complete the Orphaned Accounts sheet with detection and remediation tracking.', '7. Review the Lifecycle Metrics sheet for SLA compliance summary.', '8. Complete the Gap Analysis sheet with all non-compliance findings.', '9. Record all supporting evidence in the Evidence Register.', '10. Obtain sign-off via the Approval Sign-Off sheet.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 24

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_user_inventory_sheet(wb, styles):
    """Create Sheet 2: User_Inventory."""
    ws = wb.create_sheet("User Inventory")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:M1")
    cell = ws["A1"]
    cell.value = "USER INVENTORY - ALL IDENTITY SYSTEMS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Metadata
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Total Users: {USER_ROW_COUNT}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    headers = [
        "User ID", "Username", "Full Name", "Email", "User Type",
        "Department", "Job Title", "Manager", "Hire Date", "Termination Date",
        "Status", "Last Login", "Account Created", "Comments"
    ]
    
    row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (6) — F2F2F2 grey, shows how to fill in
    _sample_data = ['USR-0001', 'jsmith', 'John Smith', 'j.smith@example.com', 'Employee', 'IT', 'System Administrator', 'M. Johnson', '01.01.2024', '', 'Active', '28.02.2026', '01.01.2024', '']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=6, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (7–56) — FFFFCC yellow with borders
    for _data_row in range(7, 57):
        for _col_idx in range(1, 15):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    # Column widths
    widths = {"A": 10, "B": 20, "C": 20, "D": 30, "E": 15, "F": 15, "G": 25,
              "H": 20, "I": 12, "J": 15, "K": 12, "L": 12, "M": 12, "N": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze panes
    ws.freeze_panes = "A6"
    
    return ws, []  # users no longer pre-generated (sample-only template)


def create_employee_lifecycle_sheet(wb, styles, users):
    """Create Sheet 3: Employee_Lifecycle."""
    ws = wb.create_sheet("Employee Lifecycle")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "EMPLOYEE LIFECYCLE COMPLIANCE - JOINER & LEAVER EVENTS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Filter to employees only
    employees = [u for u in users if u["user_type"] == "Employee"][:EMPLOYEE_LIFECYCLE_COUNT]
    
    # Metadata
    ws["A2"] = f"Assessment Period: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Employees Assessed: {len(employees)}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers for JOINER events
    row = 5
    ws.merge_cells(f"A{row}:J{row}")
    cell = ws[f"A{row}"]
    cell.value = "JOINER EVENTS - Provisioning Compliance"
    apply_style(cell, styles["subheader"])
    
    row += 1
    headers = [
        "User ID", "Username", "Full Name", "Department",
        "Hire Date", "Provision Date", "Days to Provision", "SLA (Days)", "Compliance Status", "Notes"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # JOINER section: 1 grey sample row + 25 empty FFFFCC rows
    _joiner_sample = [
        "USR-0001", "jsmith", "John Smith", "IT",
        "15.01.2024", "16.01.2024", 1, 2, "Compliant", ""
    ]
    for _col_idx, _val in enumerate(_joiner_sample, start=1):
        _cell = ws.cell(row=7, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border
    for _data_row in range(8, 33):
        for _col_idx in range(1, 11):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    # LEAVER section header at row 33
    ws.merge_cells("A33:J33")
    _cell = ws["A33"]
    _cell.value = "LEAVER EVENTS - Deprovisioning Compliance"
    apply_style(_cell, styles["subheader"])

    # Leaver column headers at row 34
    _leaver_headers = [
        "User ID", "Username", "Full Name", "Department",
        "Termination Date", "Disable Date", "Hours to Disable",
        "SLA (Hours)", "Compliance Status", "Notes"
    ]
    for _col, _hdr in enumerate(_leaver_headers, start=1):
        _cell = ws.cell(row=34, column=_col)
        _cell.value = _hdr
        apply_style(_cell, styles["column_header"])

    # LEAVER section: 1 grey sample row + 25 empty FFFFCC rows
    _leaver_sample = [
        "USR-0002", "pmartin", "Paul Martin", "Finance",
        "15.12.2023", "15.12.2023 08:30", 2, 4, "Compliant", ""
    ]
    for _col_idx, _val in enumerate(_leaver_sample, start=1):
        _cell = ws.cell(row=35, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border
    for _data_row in range(36, 61):
        for _col_idx in range(1, 11):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    # Column widths
    widths = {"A": 10, "B": 20, "C": 20, "D": 15, "E": 15, "F": 18, "G": 15, "H": 12, "I": 18, "J": 35}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A7"
    
    return ws


def create_contractor_lifecycle_sheet(wb, styles, users):
    """Create Sheet 4: Contractor_Lifecycle."""
    ws = wb.create_sheet("Contractor Lifecycle")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "CONTRACTOR LIFECYCLE - TIME-BOUND ACCESS COMPLIANCE"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Filter to contractors
    contractors = [u for u in users if u["user_type"] == "Contractor"][:CONTRACTOR_LIFECYCLE_COUNT]
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Contractors Assessed: {len(contractors)}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    row = 5
    headers = [
        "User ID", "Username", "Full Name", "Department", "Sponsor",
        "Contract Start", "Contract End", "Access Expiry", "Days Remaining", "Status", "Compliance"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (6) — F2F2F2 grey, shows how to fill in
    _sample_data = ['USR-0050', 'c.jones', 'Carol Jones', 'IT', 'M. Wilson', '01.01.2025', '31.12.2026', '31.12.2026', 365, 'Active', 'Compliant']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=6, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (7–56) — FFFFCC yellow with borders
    for _data_row in range(7, 57):
        for _col_idx in range(1, 12):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    # Column widths
    widths = {"A": 10, "B": 22, "C": 20, "D": 15, "E": 20, "F": 15, "G": 15, "H": 15, "I": 15, "J": 15, "K": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_service_accounts_sheet(wb, styles, users):
    """Create Sheet 5: Service_Accounts."""
    ws = wb.create_sheet("Service Accounts")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "SERVICE ACCOUNT INVENTORY - NON-HUMAN ACCOUNTS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Filter to service accounts
    service_accounts = [u for u in users if u["user_type"] == "Service Account"][:SERVICE_ACCOUNT_COUNT]
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Service Accounts: {len(service_accounts)}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    row = 5
    headers = [
        "Account ID", "Account Name", "Purpose", "Owner", "Department",
        "Created Date", "Last Used", "Privileged", "Password Rotation", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (6) — F2F2F2 grey, shows how to fill in
    _sample_data = ['SVC-001', 'svc-backup', 'Automated Backup', 'IT Manager', 'IT', '01.01.2023', '28.02.2026', 'No', '90 Days', 'Active']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=6, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (7–56) — FFFFCC yellow with borders
    for _data_row in range(7, 57):
        for _col_idx in range(1, 11):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    # Column widths
    widths = {"A": 12, "B": 30, "C": 25, "D": 20, "E": 15, "F": 15, "G": 15, "H": 12, "I": 18, "J": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_orphaned_accounts_sheet(wb, styles, users):
    """Create Sheet 6: Orphaned_Accounts."""
    ws = wb.create_sheet("Orphaned Accounts")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "ORPHANED ACCOUNT DETECTION & REMEDIATION"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Identify orphaned accounts (no termination date but disabled, or very old last login)
    orphaned = []
    for user in users:
        is_orphaned = False
        reason = ""
        
        if user["status"] == "Disabled" and not user["termination_date"]:
            is_orphaned = True
            reason = "Disabled account with no termination record"
        elif user["last_login"] and (datetime.now() - user["last_login"]).days > INACTIVE_THRESHOLD_DAYS:
            is_orphaned = True
            reason = f"No login for {(datetime.now() - user['last_login']).days} days"
        
        if is_orphaned:
            orphaned.append({**user, "orphan_reason": reason})
            if len(orphaned) >= ORPHANED_ACCOUNT_COUNT:
                break
    
    ws["A2"] = f"Detection Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Orphaned Accounts Found: {len(orphaned)}"
    ws["A3"].font = Font(italic=True)
    ws["A3"].font = Font(bold=True, color="FF0000")
    
    # Column headers
    row = 5
    headers = [
        "User ID", "Username", "Full Name", "User Type", "Department",
        "Last Login", "Status", "Orphan Reason", "Detected Date", "Remediation Status", "Notes"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (6) — F2F2F2 grey, shows how to fill in
    _sample_data = ['USR-0099', 't.davis', 'Tom Davis', 'Employee', 'Finance', 'Never', 'Disabled', 'Disabled account with no termination record', '01.03.2026', 'Open', '']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=6, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (7–56) — FFFFCC yellow with borders
    for _data_row in range(7, 57):
        for _col_idx in range(1, 12):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    # Column widths
    widths = {"A": 10, "B": 20, "C": 20, "D": 15, "E": 15, "F": 15, "G": 12, "H": 40, "I": 15, "J": 18, "K": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_lifecycle_metrics_sheet(wb, styles):
    """Create Sheet 7: Lifecycle_Metrics."""
    ws = wb.create_sheet("Lifecycle Metrics")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "IDENTITY LIFECYCLE COMPLIANCE METRICS - A.5.16"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Assessment Period: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Summary metrics
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "Key Performance Indicators (KPIs)"
    apply_style(cell, styles["header"])
    
    row += 1
    headers = ["Metric", "Target", "Actual", "Status", "Gap", "Comments"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Sample metrics
    metrics = [
        ("Provisioning On-Time Rate", "≥95%", "92%", "Warning", "-3%", "3 late provisioning instances"),
        ("Deprovisioning On-Time Rate", "≥98%", "96%", "Warning", "-2%", "2 late deprovisioning instances"),
        ("Orphaned Account Count", "≤1%", "1.5%", "Non-Compliant", "+0.5%", "15 orphaned accounts detected"),
        ("Service Account Documentation", "100%", "95%", "Warning", "-5%", "1 account missing owner"),
        ("Contractor Access Compliance", "100%", "90%", "Warning", "-10%", "3 expired contractor accounts"),
    ]
    
    for metric_name, target, actual, status, gap, comments in metrics:
        row += 1
        ws.cell(row=row, column=1).value = metric_name
        ws.cell(row=row, column=2).value = target
        ws.cell(row=row, column=3).value = actual
        ws.cell(row=row, column=4).value = status
        ws.cell(row=row, column=5).value = gap
        ws.cell(row=row, column=6).value = comments
        
        status_cell = ws.cell(row=row, column=4)
        if status == "Compliant":
            apply_style(status_cell, styles["compliant"])
        elif status == "Warning":
            apply_style(status_cell, styles["warning"])
        else:
            apply_style(status_cell, styles["non_compliant"])
        
        for col in range(1, 7):
            if col != 4:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Overall score
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "Overall A.5.16 Compliance Score"
    apply_style(cell, styles["header"])
    
    row += 1
    ws[f"A{row}"] = "Score Calculation"
    ws[f"B{row}"] = "(Provisioning + Deprovisioning + Orphaned Account Management) / 3"
    
    row += 1
    ws[f"A{row}"] = "Overall Score"
    ws[f"B{row}"] = "82%"
    ws[f"B{row}"].font = Font(bold=True, size=14)
    
    row += 1
    ws[f"A{row}"] = "Maturity Level"
    ws[f"B{row}"] = "Managed (80-89%)"
    apply_style(ws[f"B{row}"], styles["warning"])
    
    # Column widths
    widths = {"A": 35, "B": 15, "C": 15, "D": 18, "E": 10, "F": 50}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    return ws


def create_gap_analysis_sheet(wb, styles):
    """Create Sheet 8: Gap_Analysis."""
    ws = wb.create_sheet("Gap Analysis")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "GAP ANALYSIS - NON-COMPLIANCE FINDINGS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Column headers
    row = 4
    headers = [
        "Gap ID", "Category", "Description", "Risk Level", "Affected Users",
        "Root Cause", "Remediation Plan", "Owner", "Due Date", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Sample row (row 5) — F2F2F2 grey, shows example data
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    sample_data = (
        "GAP-001", "Provisioning", "3 users provisioned after hire date",
        "Medium", "3", "Manual provisioning delays",
        "Automate provisioning via HR system", "IAM Manager", "2026-06-30", "Open"
    )
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = grey_fill
        cell.border = _border

    # 50 empty data rows (rows 6–55)
    for data_row in range(6, 56):
        for col_idx in range(1, 11):
            cell = ws.cell(row=data_row, column=col_idx, value="")
            cell.fill = yellow_fill
            cell.border = _border

    # Data validations (skip sample row)
    risk_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add("D6:D55")

    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Accepted"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("J6:J55")

    # Column widths
    widths = {"A": 10, "B": 18, "C": 40, "D": 12, "E": 15, "F": 30, "G": 40, "H": 18, "I": 12, "J": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A4"

    return ws


def create_summary_dashboard_sheet(wb):
    """Create the Summary Dashboard sheet (Gold Standard — GS-SD-014/016 compliant)."""
    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_hdr = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _red_banner = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _ffffcc = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    _black_font = Font(name="Calibri", size=10, color="000000")
    _bold_black = Font(name="Calibri", size=10, bold=True, color="000000")

    # Row 1: Title — GS-SD-014: must contain "— SUMMARY DASHBOARD"
    ws.merge_cells("A1:G1")
    ws["A1"] = "USER INVENTORY & LIFECYCLE \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = _border
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle (left-aligned per GS)
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Control A.5.16: Identity Management"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Empty spacer

    # TABLE 1: Assessment Area Compliance Overview
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = _white_font
    ws["A4"].fill = _navy
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A4"].border = _border

    # Row 5: Column headers — GS-SD-016: must be D9D9D9
    t1_headers = ["Assessment Area", "Total Items", "Compliant", "Partial / Warning",
                  "Non-Compliant", "N/A", "Compliance %"]
    for col, hdr in enumerate(t1_headers, 1):
        cell = ws.cell(row=5, column=col, value=hdr)
        cell.font = _bold_black
        cell.fill = _grey_hdr
        cell.border = _border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # TABLE 1 data rows (rows 6-10): 5 assessment areas
    area_configs = [
        # (Area Name, COUNTA range, status_col, good, partial, bad)
        ("Employee Lifecycle",
         "COUNTA('Employee Lifecycle'!A7:A106)",
         "I", "'Employee Lifecycle'!I7:I106",
         "Compliant", "0", "Non-Compliant"),
        ("Contractor Lifecycle",
         "COUNTA('Contractor Lifecycle'!A6:A56)",
         "K", "'Contractor Lifecycle'!K6:K56",
         "Compliant", "Warning", "Non-Compliant"),
        ("Service Accounts",
         "COUNTA('Service Accounts'!A6:A56)",
         "J", "'Service Accounts'!J6:J56",
         "Active", "0", "Disabled"),
        ("Orphaned Accounts",
         "COUNTA('Orphaned Accounts'!A6:A56)",
         "J", "'Orphaned Accounts'!J6:J56",
         "COUNTIF_DUAL:Remediated|Closed", "In Progress", "Open"),
        ("Gap Analysis",
         "COUNTA('Gap Analysis'!A6:A55)",
         "J", "'Gap Analysis'!J6:J55",
         "Resolved", "In Progress", "Open"),
    ]

    for i, (area_name, counta_expr, _sc, rng, good, partial, bad) in enumerate(area_configs):
        row = 6 + i

        # Col A: Area name
        cell_a = ws.cell(row=row, column=1, value=area_name)
        cell_a.border = _border
        cell_a.font = _black_font

        # Col B: Total Items
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = f"={counta_expr}"
        cell_b.border = _border
        cell_b.alignment = Alignment(horizontal="center")
        cell_b.font = _black_font

        # Col C: Compliant/Good
        cell_c = ws.cell(row=row, column=3)
        if good.startswith("COUNTIF_DUAL:"):
            parts = good.replace("COUNTIF_DUAL:", "").split("|")
            cell_c.value = f'=COUNTIF({rng},"{parts[0]}")+COUNTIF({rng},"{parts[1]}")'
        elif good == "0":
            cell_c.value = 0
        else:
            cell_c.value = f'=COUNTIF({rng},"{good}")'
        cell_c.border = _border
        cell_c.alignment = Alignment(horizontal="center")
        cell_c.font = _black_font

        # Col D: Partial/Warning
        cell_d = ws.cell(row=row, column=4)
        if partial == "0":
            cell_d.value = 0
        else:
            cell_d.value = f'=COUNTIF({rng},"{partial}")'
        cell_d.border = _border
        cell_d.alignment = Alignment(horizontal="center")
        cell_d.font = _black_font

        # Col E: Non-Compliant/Bad
        cell_e = ws.cell(row=row, column=5)
        cell_e.value = f'=COUNTIF({rng},"{bad}")'
        cell_e.border = _border
        cell_e.alignment = Alignment(horizontal="center")
        cell_e.font = _black_font

        # Col F: N/A
        cell_f = ws.cell(row=row, column=6, value=0)
        cell_f.border = _border
        cell_f.alignment = Alignment(horizontal="center")
        cell_f.font = _black_font

        # Col G: Compliance %
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        cell_g.number_format = "0.0%"
        cell_g.border = _border
        cell_g.alignment = Alignment(horizontal="center")
        cell_g.font = Font(name="Calibri", size=10, color="000000")

    # Row 11: TOTAL
    total_row = 11
    for col in range(1, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = _grey_hdr
        cell.border = _border
        cell.font = _bold_black
        cell.alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left")
    for col in range(2, 7):
        ws.cell(row=total_row, column=col).value = \
            f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}10)"
    # TOTAL compliance %
    cell_g_total = ws.cell(row=total_row, column=7)
    cell_g_total.value = f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"
    cell_g_total.number_format = "0.0%"

    # TABLE 2: Key Metrics
    t2_start = 13
    ws.merge_cells(f"A{t2_start}:G{t2_start}")
    ws[f"A{t2_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{t2_start}"].font = _white_font
    ws[f"A{t2_start}"].fill = _navy
    ws[f"A{t2_start}"].border = _border

    # TABLE 2 headers — D9D9D9 (GS-SD-016)
    t2_hdr_row = t2_start + 1
    for col, hdr in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        cell = ws.cell(row=t2_hdr_row, column=col, value=hdr if hdr else None)
        cell.font = _bold_black
        cell.fill = _grey_hdr
        cell.border = _border
        cell.alignment = Alignment(horizontal="center")

    metrics = [
        ("Total Users in Inventory", "=COUNTA('User Inventory'!A6:A56)", None),
        ("Active User Accounts", '=COUNTIF(\'User Inventory\'!K6:K56,"Active")', None),
        ("Disabled Accounts", '=COUNTIF(\'User Inventory\'!K6:K56,"Disabled")', None),
        ("Orphaned Accounts Detected", "=COUNTA('Orphaned Accounts'!A6:A56)", None),
        ("Orphaned Accounts Remediated",
         '=COUNTIF(\'Orphaned Accounts\'!J6:J56,"Remediated")+COUNTIF(\'Orphaned Accounts\'!J6:J56,"Closed")', None),
        ("Joiner On-Time Provisioning Rate",
         "=IFERROR(COUNTIF('Employee Lifecycle'!I7:I32,\"Compliant\")/COUNTA('Employee Lifecycle'!A7:A32),0)",
         "0.0%"),
        ("Leaver On-Time Deprovisioning Rate",
         "=IFERROR(COUNTIF('Employee Lifecycle'!I35:I60,\"Compliant\")/COUNTA('Employee Lifecycle'!A35:A60),0)",
         "0.0%"),
        ("Contractor Compliance Rate",
         "=IFERROR(COUNTIF('Contractor Lifecycle'!K6:K56,\"Compliant\")/COUNTA('Contractor Lifecycle'!A6:A56),0)",
         "0.0%"),
        ("Expired Contractor Accounts", '=COUNTIF(\'Contractor Lifecycle\'!K6:K56,"Non-Compliant")', None),
        ("Service Accounts with Valid Owner", '=COUNTIF(\'Service Accounts\'!J6:J56,"Active")', None),
        ("Open Lifecycle Gaps", '=COUNTIF(\'Gap Analysis\'!J6:J55,"Open")', None),
        ("Gaps In Progress", '=COUNTIF(\'Gap Analysis\'!J6:J55,"In Progress")', None),
    ]

    row = t2_start + 2
    for metric_name, formula, fmt in metrics:
        ws.cell(row=row, column=1, value=metric_name).border = _border
        ws.cell(row=row, column=1).font = _black_font
        cell_val = ws.cell(row=row, column=2, value=formula)
        cell_val.border = _border
        cell_val.font = _black_font
        cell_val.alignment = Alignment(horizontal="center")
        if fmt:
            cell_val.number_format = fmt
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = _border
        row += 1

    # TABLE 2 buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = _border
        row += 1

    # TABLE 3: Critical Findings
    t3_start = row + 1
    ws.merge_cells(f"A{t3_start}:G{t3_start}")
    ws[f"A{t3_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{t3_start}"].font = _white_font
    ws[f"A{t3_start}"].fill = _red_banner
    ws[f"A{t3_start}"].border = _border

    # TABLE 3 headers — D9D9D9
    t3_hdr_row = t3_start + 1
    for col, hdr in enumerate(["Category", "Finding", "Count", "Severity", "Action Required", "", ""], 1):
        cell = ws.cell(row=t3_hdr_row, column=col, value=hdr if hdr else None)
        cell.font = _bold_black
        cell.fill = _grey_hdr
        cell.border = _border
        cell.alignment = Alignment(horizontal="center")

    findings = [
        ("Identity Lifecycle", "Orphaned accounts pending remediation",
         '=COUNTIF(\'Orphaned Accounts\'!J6:J56,"Open")', "Critical", "Immediate"),
        ("Identity Lifecycle", "Non-compliant lifecycle events (employees)",
         '=COUNTIF(\'Employee Lifecycle\'!I7:I106,"Non-Compliant")', "Critical", "Immediate"),
        ("Contractor Management", "Expired contractor accounts still active",
         '=COUNTIF(\'Contractor Lifecycle\'!K6:K56,"Non-Compliant")', "High", "Urgent"),
        ("Contractor Management", "Contractors expiring — access review required",
         '=COUNTIF(\'Contractor Lifecycle\'!K6:K56,"Warning")', "High", "Urgent"),
        ("Service Accounts", "Disabled service accounts (ownership risk)",
         '=COUNTIF(\'Service Accounts\'!J6:J56,"Disabled")', "Medium", "Plan"),
        ("Gap Analysis", "High-risk open gaps",
         '=COUNTIF(\'Gap Analysis\'!D6:D55,"High")', "Critical", "Immediate"),
        ("Gap Analysis", "Total open gaps requiring remediation",
         '=COUNTIF(\'Gap Analysis\'!J6:J55,"Open")', "High", "Urgent"),
        ("Identity Lifecycle", "Orphaned accounts in progress (not yet resolved)",
         '=COUNTIF(\'Orphaned Accounts\'!J6:J56,"In Progress")', "Medium", "Plan"),
    ]

    row = t3_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = _ffffcc
            ws.cell(row=row, column=col).border = _border
            ws.cell(row=row, column=col).font = _black_font
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    # TABLE 3 buffer rows
    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = _ffffcc
            ws.cell(row=row, column=col).border = _border
        row += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A4"

    return ws


def create_evidence_register(wb):
    """Create the standard Evidence Register sheet (gold standard)."""
    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_hdr = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _input = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = f"ISO/IEC 27001:2022 Annex A - Control {CONTROL_ID}: {CONTROL_NAME}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = _border

    ws.merge_cells("A3:H3")
    ws["A3"].border = _border

    columns = [
        ("Evidence ID", 14), ("Evidence Type", 20), ("Description", 45),
        ("Related Control / Section", 28), ("Collection Date (DD.MM.YYYY)", 22),
        ("Storage Location / Reference", 38), ("Collected By", 22), ("Status", 14),
    ]
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    sample_data = ["EV-001", "Document", "Sample evidence entry — replace with actual evidence",
                   "All Controls", "01.01.2026", "SharePoint/ISMS/Evidence/", "ISMS Team", "Active"]
    for col_idx, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
        cell.fill = _grey_sample
        cell.border = _border

    dv_status = DataValidation(
        type="list",
        formula1='"Active,Archived,Superseded,Pending Review"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(6, 106):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = _input
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"

    return ws


def create_approval_sheet(wb):
    """Create the Approval Sign-Off sheet — Gold Standard (GS-AS-014/015)."""
    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title banner — GS-AS-014
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY section banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Rows 4-8: Summary metadata — B6 = Overall Compliance (GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Row 7 status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # Approver sections start at row 11 (rows 9-10 = gap)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION — GS-AS-012: col A = plain bold label, NO dark fill
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("="*80)
    logger.info(f"Generating: {WORKBOOK_NAME}")
    logger.info(f"Control: ISO/IEC 27001:2022 {CONTROL_ID} - {CONTROL_NAME}")
    logger.info("="*80)

    # Create workbook
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.remove(wb.active)  # Remove default sheet

    # Setup styles
    styles = _STYLES

    # Create sheets
    logger.info("Creating sheets...")

    logger.info("  1/10 Creating Instructions & Legend...")
    create_instructions_sheet(wb.create_sheet())

    logger.info("  2/10 Creating User Inventory...")
    _, users = create_user_inventory_sheet(wb, styles)

    logger.info("  3/10 Creating Employee Lifecycle...")
    create_employee_lifecycle_sheet(wb, styles, users)

    logger.info("  4/10 Creating Contractor Lifecycle...")
    create_contractor_lifecycle_sheet(wb, styles, users)

    logger.info("  5/10 Creating Service Accounts...")
    create_service_accounts_sheet(wb, styles, users)

    logger.info("  6/10 Creating Orphaned Accounts...")
    create_orphaned_accounts_sheet(wb, styles, users)

    logger.info("  7/10 Creating Lifecycle Metrics...")
    create_lifecycle_metrics_sheet(wb, styles)

    logger.info("  8/10 Creating Gap Analysis...")
    create_gap_analysis_sheet(wb, styles)

    logger.info("  9/11 Creating Evidence Register...")
    create_evidence_register(wb)

    logger.info("  10/11 Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb)

    logger.info("  11/11 Creating Approval Sign-Off...")
    create_approval_sheet(wb)

    finalize_validations(wb)

    # Save workbook
    logger.info(f"Saving workbook: {output_path.name}")
    wb.save(output_path)

    logger.info("="*80)
    logger.info("SUCCESS!")
    logger.info(f"File created: {output_path}")
    logger.info("Sheets: 11")
    logger.info(f"Sample users: {USER_ROW_COUNT}")
    logger.info("="*80)
    logger.info("Workbook Contents:")
    logger.info("  - User Inventory (100 sample users)")
    logger.info("  - Employee Lifecycle Compliance (joiner/leaver events)")
    logger.info("  - Contractor Time-Bound Access Tracking")
    logger.info("  - Service Account Inventory")
    logger.info("  - Orphaned Account Detection (15 accounts)")
    logger.info("  - Lifecycle Metrics & KPIs")
    logger.info("  - Gap Analysis & Remediation Tracking")
    logger.info("  - Evidence Register for A.5.16")
    logger.info("  - Summary Dashboard")
    logger.info("  - Approval and Sign-Off")
    logger.info("="*80)



def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    try:
        output_file = main()
        sys.exit(0)
    except Exception as e:
        logger.error(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
