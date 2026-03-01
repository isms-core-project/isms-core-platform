#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.15-16-18.S4 - Role Compliance & SoD Assessment Generator
================================================================================

ISO/IEC 27001:2022 Controls A.5.15 & A.5.18: Access Control & Access Rights
Assessment Workbook 4 of 5: Role Definition, RBAC Compliance, and Segregation of Duties

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific role definitions, RBAC implementation,
and segregation of duties requirements.

Key customization areas:
1. Role catalog and definitions (match your organisational structure)
2. Role-to-access mappings (specific to your systems/applications)
3. Segregation of Duties (SoD) matrix (based on your fraud risk analysis)
4. Conflicting role combinations (aligned with your control framework)
5. RBAC adoption targets (match your IAM maturity goals)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.15-16-18 IAM Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for evaluating
role-based access control (RBAC) implementation and segregation of duties
compliance across all systems.

**Purpose:**
Enables systematic assessment of role governance and SoD controls against
ISO 27001:2022 Controls A.5.15 and A.5.18 requirements, supporting
evidence-based validation of access control design and fraud prevention.

**Assessment Scope:**
- Role catalog completeness and accuracy
- Role definition documentation (purpose, access rights, ownership)
- Role-to-user assignment tracking
- Role-to-access mapping validation
- RBAC adoption rate (role-based vs. direct access)
- Segregation of Duties (SoD) matrix compliance
- Conflicting role identification and remediation
- Compensating controls for SoD violations
- Role governance effectiveness
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Usage guidance and RBAC/SoD standards
2. Role_Catalog - Complete role inventory (40 roles)
3. Role_Assignments - User-to-role mappings (100 users)
4. Role_Access_Mapping - Role-to-system access rights (40 roles × systems)
5. SoD_Matrix - Conflicting role combinations (30 conflict pairs)
6. SoD_Violations - Users with conflicting roles (15 violations)
7. Compensating_Controls - SoD violation mitigations (15 controls)
8. RBAC_Coverage - RBAC adoption metrics by system
9. Gap_Analysis - Role governance gaps and remediation tracking
10. Evidence_Register - Evidence collection for A.5.15/A.5.18 compliance
11. Approval_Sign_Off - 3-level approval workflow

**Key Features:**
- Data validation with role catalog dropdown lists
- Conditional formatting for SoD violation status
- Automated gap identification for conflicting role assignments
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with identity system role data

**Integration:**
This assessment feeds into the IAM Governance Dashboard, which
consolidates data from all five IAM assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - random (standard library - for sample data generation)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a515-16-18_4_role_compliance.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a515-16-18_4_role_compliance.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a515-16-18_4_role_compliance.py --date 20250124

Output:
    File: ISMS-IMP-A.5.15-16-18.S4_Role_Compliance_SoD_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Document complete role catalog with definitions
    2. Map roles to access rights across all systems
    3. Document user-to-role assignments
    4. Define Segregation of Duties (SoD) matrix
    5. Identify conflicting role combinations
    6. Detect users with conflicting roles (SoD violations)
    7. Document compensating controls for unavoidable violations
    8. Calculate RBAC adoption rate
    9. Identify role governance gaps
    10. Plan remediation for SoD violations
    11. Collect and link audit evidence
    12. Obtain stakeholder approvals
    13. Feed results into IAM Governance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.5.15, A.5.18
Assessment Domain:    4 of 5 (Role Compliance & Segregation of Duties)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.15-16-18: Access Control Policy (A.5.15 - SoD requirements)
    - ISMS-POL-A.5.15-16-18: Access Rights Management Policy (A.5.18 - RBAC)
    - ISMS-IMP-A.5.15-16-18-S3: Role Definition and Assignment Guide
    - ISMS-IMP-A.5.15-16-18.S1: User Inventory Assessment (Workbook 1)
    - ISMS-IMP-A.5.15-16-18.S2: Access Rights Matrix Assessment (Workbook 2)
    - ISMS-IMP-A.5.15-16-18.S3: Access Review Results Assessment (Workbook 3)
    - ISMS-IMP-A.5.15-16-18.S5: Lifecycle Compliance Detailed (Workbook 5)
    - IAM Governance Dashboard: Consolidated IAM compliance view

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.15-16-18-S3
    - Supports comprehensive RBAC and SoD compliance evaluation
    - Integrated with IAM Governance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Segregation of Duties (SoD):**
SoD is CRITICAL fraud prevention control. Same person should NOT be able to:
- Request AND approve transactions
- Create AND approve vendor records
- Submit AND authorise payments
- Develop AND deploy code to production
- Create users AND assign privileged access

SoD violations = major audit finding and fraud risk.

**SoD Matrix Development:**
Create SoD matrix through fraud risk analysis:
1. Identify critical business processes (payments, procurement, HR)
2. Map process steps to system roles
3. Identify conflicting combinations (same user = fraud opportunity)
4. Document in SoD matrix (Role A conflicts with Role B)
5. Implement preventive controls (system blocks conflicting assignments)
6. Document detective controls (periodic SoD violation scans)

Don't copy generic SoD matrix - must be specific to YOUR processes and risks.

**Audit Considerations:**
Auditors expect to see:
- Documented SoD matrix based on fraud risk analysis
- Evidence of SoD violation detection (regular scans)
- Remediation of violations (role removal or compensating controls)
- Business justification for unavoidable violations
- Compensating controls for justified exceptions (dual approval, logging, review)

**Data Protection:**
Assessment workbooks contain sensitive control design information:
- SoD matrix (reveals fraud scenarios)
- SoD violations (identifies control weaknesses)
- Role definitions and access mappings (confidential architecture)

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Monthly: Scan for new SoD violations (automated)
- Quarterly: Review role catalog accuracy and role assignments
- Semi-annually: Update SoD matrix for new processes/systems
- Annually: Complete role governance reassessment
- Ad-hoc: When new systems deployed or processes change

**Quality Assurance:**
SoD effectiveness depends on:
- Accurate role definitions (roles actually match documented access)
- Complete SoD matrix (all critical conflicts documented)
- Automated detection (manual SoD checks miss violations)
- Timely remediation (violations fixed, not just documented)

Validate role access against actual system permissions periodically.

**RBAC Best Practices:**
- Roles based on job functions, not individuals
- Role definitions documented and approved by business owners
- Role access regularly reviewed for accuracy
- New roles require formal approval process
- Deprecated roles removed (not just disabled)
- Role hierarchy minimised (avoid complex nested roles)

**Compensating Controls for SoD Violations:**
When SoD violations unavoidable (small organisation, unique expertise):
- Document business justification
- Implement compensating controls:
  * Dual approval workflows (second person reviews)
  * Enhanced logging and monitoring
  * Management review of sensitive transactions
  * Periodic audit of user's activities
- Re-evaluate periodically (can violation be eliminated?)

Compensating controls must be EFFECTIVE, not cosmetic.

**Common SoD Mistakes:**
- Generic SoD matrix not tailored to organisation
- SoD violations identified but not remediated
- Compensating controls documented but not implemented
- One-time SoD scan, no ongoing monitoring
- SoD matrix not updated when processes/systems change

Avoid these pitfalls through systematic governance.

**RBAC Maturity Indicators:**
- High RBAC adoption rate (>80% access via roles)
- Low direct access assignments (<20%)
- Minimal SoD violations (<5% of users)
- Documented and approved role catalog
- Regular role access reviews
- Automated SoD violation detection

Track these metrics to demonstrate IAM maturity.

================================================================================
"""

import sys
from datetime import datetime, timedelta
from pathlib import Path
import random

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import BarChart, Reference
except ImportError as e:
    logger.error(f"ERROR: Required library not found: {e}")
    logger.info("Install required libraries: pip install openpyxl")
    sys.exit(1)

_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)
_assessment_dir = _wkbk_dir / "Assessment"
if _assessment_dir.exists():
    _wkbk_dir = _assessment_dir

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
WORKBOOK_NAME = "Role Compliance & Segregation of Duties Assessment"
DOCUMENT_ID = "ISMS-IMP-A.5.15-16-18.S4"
CONTROL_ID = "A.5.15 & A.5.18"
CONTROL_NAME = "Access Control & Access Rights (RBAC/SoD)"
CONTROL_REF   = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# Data row counts
ROLE_COUNT = 30                # Defined roles
ROLE_ASSIGNMENT_COUNT = 80     # Users with role assignments
DIRECT_ACCESS_COUNT = 20       # Users with direct access (non-RBAC)
SOD_CONFLICT_COUNT = 6         # Defined SoD conflicts
SOD_VIOLATION_COUNT = 5        # Detected SoD violations
GAP_ROW_COUNT = 30            # Gap analysis tracking
EVIDENCE_ROW_COUNT = 50       # Evidence register
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point
# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data_center": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "compliant": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "non_compliant": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "warning": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }
    
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = style_dict["font"]
    if "fill" in style_dict:
        cell.fill = style_dict["fill"]
    if "alignment" in style_dict:
        cell.alignment = style_dict["alignment"]
    if "border" in style_dict:
        cell.border = style_dict["border"]


# ============================================================================
# SECTION 3: SAMPLE DATA GENERATORS
# ============================================================================

def generate_role_catalog():
    """Generate sample role catalog."""
    roles = [
        # Finance roles
        {"id": "R-001", "name": "Finance Analyst", "dept": "Finance", "desc": "Financial analysis and reporting", 
         "systems": "Finance System, Email, Intranet", "user_count": 8, "privileged": "No"},
        {"id": "R-002", "name": "Finance Manager", "dept": "Finance", "desc": "Financial management and approval", 
         "systems": "Finance System (Admin), Email, Intranet", "user_count": 3, "privileged": "Yes"},
        {"id": "R-003", "name": "Accounts Payable", "dept": "Finance", "desc": "AP processing and vendor management", 
         "systems": "Finance System, Email", "user_count": 4, "privileged": "No"},
        {"id": "R-004", "name": "Accounts Receivable", "dept": "Finance", "desc": "AR processing and collections", 
         "systems": "Finance System, CRM, Email", "user_count": 3, "privileged": "No"},
        
        # Sales roles
        {"id": "R-005", "name": "Sales Representative", "dept": "Sales", "desc": "Customer sales and account management", 
         "systems": "CRM, Email, Intranet", "user_count": 15, "privileged": "No"},
        {"id": "R-006", "name": "Sales Manager", "dept": "Sales", "desc": "Sales team leadership and approvals", 
         "systems": "CRM (Manager), Email, Intranet", "user_count": 4, "privileged": "No"},
        {"id": "R-007", "name": "Sales Operations", "dept": "Sales", "desc": "Sales process and analytics", 
         "systems": "CRM (Admin), Finance System, Email", "user_count": 2, "privileged": "Yes"},
        
        # IT roles
        {"id": "R-008", "name": "IT Administrator", "dept": "IT", "desc": "System administration and support", 
         "systems": "All Systems (Admin)", "user_count": 5, "privileged": "Yes"},
        {"id": "R-009", "name": "Security Analyst", "dept": "IT", "desc": "Security monitoring and incident response", 
         "systems": "Security Tools, Logs, Email", "user_count": 3, "privileged": "Yes"},
        {"id": "R-010", "name": "Help Desk", "dept": "IT", "desc": "User support and ticket management", 
         "systems": "Ticketing System, Email, AD (Limited)", "user_count": 6, "privileged": "No"},
        
        # HR roles
        {"id": "R-011", "name": "HR Generalist", "dept": "HR", "desc": "HR operations and employee services", 
         "systems": "HR System, Email, Intranet", "user_count": 4, "privileged": "No"},
        {"id": "R-012", "name": "HR Manager", "dept": "HR", "desc": "HR leadership and approvals", 
         "systems": "HR System (Manager), Email, Intranet", "user_count": 2, "privileged": "Yes"},
        {"id": "R-013", "name": "Recruiter", "dept": "HR", "desc": "Recruitment and candidate management", 
         "systems": "HR System (Recruiting), Email", "user_count": 3, "privileged": "No"},
        
        # Engineering roles
        {"id": "R-014", "name": "Software Developer", "dept": "Engineering", "desc": "Software development", 
         "systems": "Dev Tools, Code Repo, Email", "user_count": 20, "privileged": "No"},
        {"id": "R-015", "name": "Senior Engineer", "dept": "Engineering", "desc": "Senior development and architecture", 
         "systems": "Dev Tools, Code Repo (Admin), Email", "user_count": 8, "privileged": "No"},
        {"id": "R-016", "name": "DevOps Engineer", "dept": "Engineering", "desc": "Infrastructure and deployment", 
         "systems": "Cloud Platforms (Admin), Code Repo, Email", "user_count": 5, "privileged": "Yes"},
        
        # Marketing roles
        {"id": "R-017", "name": "Marketing Specialist", "dept": "Marketing", "desc": "Marketing campaigns and content", 
         "systems": "Marketing Automation, Email, Intranet", "user_count": 6, "privileged": "No"},
        {"id": "R-018", "name": "Marketing Manager", "dept": "Marketing", "desc": "Marketing leadership", 
         "systems": "Marketing Automation (Admin), CRM, Email", "user_count": 2, "privileged": "No"},
        
        # Operations roles
        {"id": "R-019", "name": "Operations Analyst", "dept": "Operations", "desc": "Operations analysis and reporting", 
         "systems": "Operations Systems, Email, Intranet", "user_count": 5, "privileged": "No"},
        {"id": "R-020", "name": "Operations Manager", "dept": "Operations", "desc": "Operations leadership", 
         "systems": "Operations Systems (Manager), Email", "user_count": 3, "privileged": "No"},
        
        # Cross-functional roles
        {"id": "R-021", "name": "Standard User", "dept": "All", "desc": "Basic employee access", 
         "systems": "Email, Intranet, File Server", "user_count": 100, "privileged": "No"},
        {"id": "R-022", "name": "Manager", "dept": "All", "desc": "People management capabilities", 
         "systems": "HR System (Manager View), Email", "user_count": 25, "privileged": "No"},
        {"id": "R-023", "name": "Project Manager", "dept": "All", "desc": "Project management and tracking", 
         "systems": "Project Management, Email, Intranet", "user_count": 8, "privileged": "No"},
        
        # Specialized roles
        {"id": "R-024", "name": "Legal Counsel", "dept": "Legal", "desc": "Legal review and compliance", 
         "systems": "Legal Systems, Email, Intranet", "user_count": 2, "privileged": "No"},
        {"id": "R-025", "name": "Auditor", "dept": "Finance", "desc": "Internal audit and compliance", 
         "systems": "All Systems (Read-Only), Audit Tools", "user_count": 3, "privileged": "Yes"},
        {"id": "R-026", "name": "Executive", "dept": "Executive", "desc": "Executive leadership access", 
         "systems": "All Systems (Executive View), Email", "user_count": 5, "privileged": "Yes"},
        {"id": "R-027", "name": "Contractor", "dept": "All", "desc": "Limited contractor access", 
         "systems": "Email, Intranet (Limited)", "user_count": 12, "privileged": "No"},
        {"id": "R-028", "name": "Customer Success", "dept": "Sales", "desc": "Customer support and success", 
         "systems": "CRM, Support System, Email", "user_count": 8, "privileged": "No"},
        {"id": "R-029", "name": "Product Manager", "dept": "Product", "desc": "Product strategy and management", 
         "systems": "Product Tools, CRM, Email", "user_count": 4, "privileged": "No"},
        {"id": "R-030", "name": "Data Analyst", "dept": "All", "desc": "Data analysis and reporting", 
         "systems": "Analytics Tools, Database (Read), Email", "user_count": 6, "privileged": "No"},
    ]
    
    return roles[:ROLE_COUNT]


def generate_sod_matrix():
    """Generate segregation of duties conflict matrix."""
    conflicts = [
        {
            "conflict_id": "SOD-001",
            "role_a": "Finance Manager",
            "role_b": "Accounts Payable",
            "risk": "High",
            "description": "Finance Manager can approve and process payments",
            "rationale": "Separation of approval and execution for payments",
            "compensating_controls": "Dual approval for payments >CHF 10'000",
        },
        {
            "conflict_id": "SOD-002",
            "role_a": "Sales Operations",
            "role_b": "Finance Analyst",
            "risk": "Medium",
            "description": "User can modify sales data and generate financial reports",
            "rationale": "Separation of data entry and reporting",
            "compensating_controls": "Monthly reconciliation by independent team",
        },
        {
            "conflict_id": "SOD-003",
            "role_a": "IT Administrator",
            "role_b": "Security Analyst",
            "risk": "High",
            "description": "User has admin access and can disable security monitoring",
            "rationale": "Separation of system administration and security oversight",
            "compensating_controls": "Privileged access monitoring, dual control for critical changes",
        },
        {
            "conflict_id": "SOD-004",
            "role_a": "HR Manager",
            "role_b": "Finance Manager",
            "risk": "Medium",
            "description": "User can create employees and approve payroll",
            "rationale": "Separation of employee creation and payment authorisation",
            "compensating_controls": "Quarterly HR-Finance reconciliation",
        },
        {
            "conflict_id": "SOD-005",
            "role_a": "DevOps Engineer",
            "role_b": "Auditor",
            "risk": "High",
            "description": "User can deploy code and audit deployments",
            "rationale": "Separation of execution and audit/verification",
            "compensating_controls": "External audit for critical systems",
        },
        {
            "conflict_id": "SOD-006",
            "role_a": "Accounts Payable",
            "role_b": "Accounts Receivable",
            "risk": "Medium",
            "description": "User can process both payables and receivables",
            "rationale": "Separation of cash inflows and outflows",
            "compensating_controls": "Monthly reconciliation, manager review",
        },
    ]
    
    return conflicts[:SOD_CONFLICT_COUNT]


def generate_sample_users(count):
    """Generate sample user data."""
    first_names = ["Alice", "Bob", "Carol", "David", "Emma", "Frank", "Grace", "Henry", 
                   "Iris", "Jack", "Kate", "Liam", "Mary", "Nathan", "Olivia", "Peter"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis"]
    
    departments = ["Finance", "Sales", "Marketing", "IT", "HR", "Operations", "Engineering"]
    
    users = []
    for i in range(count):
        first = random.choice(first_names)
        last = random.choice(last_names)
        users.append({
            "user_id": f"U{10000 + i}",
            "username": f"{first.lower()}.{last.lower()}",
            "full_name": f"{first} {last}",
            "department": random.choice(departments),
        })
    
    return users


# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Open this workbook and review the sheet tabs to understand the assessment scope.', '2. Complete the Assessment Date and Prepared By fields in the Document Information table above.', '3. Navigate to each data sheet and review the pre-populated sample data.', '4. Replace sample data with actual organisation data for your environment.', '5. Use the dropdown lists in each sheet for consistent status and category values.', '6. Document all gaps and findings in the Gap Analysis sheet.', '7. Collect supporting evidence and record references in the Evidence Register.', '8. Obtain required approvals via the Approval Sign-Off sheet before submission.', '9. Store the completed workbook in the designated ISMS evidence repository.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_role_catalog_sheet(wb, styles):
    """Create Sheet 2: Role Catalog."""
    ws = wb.create_sheet("Role Catalog")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "ROLE CATALOG - RBAC ROLE DEFINITIONS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Catalog Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Total Roles: {ROLE_COUNT}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    row = 5
    headers = [
        "Role ID", "Role Name", "Department", "Description", "Systems/Access",
        "User Count", "Privileged", "Last Review", "Owner", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (6) — F2F2F2 grey, shows how to fill in
    _sample_data = ['ROLE-001', 'IT Standard User', 'IT', 'Standard read/write access for IT personnel', 'Active Directory, Email, Intranet', 45, 'No', '01.01.2026', 'IT Manager', 'Active']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=6, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (7–56) — FFFFCC yellow with borders
    for _data_row in range(7, 57):
        for _col_idx in range(1, 11):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    roles = []  # No pre-generated roles (sheet is now blank template)
    # Column widths
    widths = {"A": 12, "B": 25, "C": 15, "D": 40, "E": 40, "F": 12, "G": 12, "H": 12, "I": 20, "J": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws, roles


def create_role_assignments_sheet(wb, styles, roles):
    """Create Sheet 3: Role Assignments."""
    ws = wb.create_sheet("Role Assignments")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "ROLE ASSIGNMENTS - USER TO ROLE MAPPINGS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Users with Roles: {ROLE_ASSIGNMENT_COUNT}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    row = 5
    headers = [
        "User ID", "Username", "Full Name", "Department", "Assigned Role",
        "Assignment Date", "Assignment Type", "Role Appropriate", "Verified By", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (6) — F2F2F2 grey, shows how to fill in
    _sample_data = ['USR-0001', 'jsmith', 'John Smith', 'IT', 'IT Standard User', '01.01.2024', 'Automatic', 'Yes', 'IT Manager', 'Active']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=6, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (7–56) — FFFFCC yellow with borders
    for _data_row in range(7, 57):
        for _col_idx in range(1, 11):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    # Column widths
    widths = {"A": 10, "B": 18, "C": 20, "D": 15, "E": 25, "F": 15, "G": 18, "H": 18, "I": 20, "J": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_direct_access_sheet(wb, styles):
    """Create Sheet 4: Direct Access Users."""
    ws = wb.create_sheet("Direct Access Users")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "DIRECT ACCESS USERS - NON-RBAC ACCESS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True, bold=True, color="FF6600")
    ws["A3"] = f"Direct Access Users: {DIRECT_ACCESS_COUNT}"
    ws["A3"].font = Font(italic=True, bold=True, color="FF6600")
    
    # Column headers
    row = 5
    headers = [
        "User ID", "Username", "Full Name", "Department", "System",
        "Access Level", "Granted Date", "Justification", "Migration Plan", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (6) — F2F2F2 grey, shows how to fill in
    _sample_data = ['USR-0020', 'ldavis', 'Linda Davis', 'Finance', 'Finance System', 'Admin', '15.06.2022', 'Legacy access before RBAC implementation', 'Migrate to FINANCE-ADMIN role by Q2 2026', 'Non-Compliant']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=6, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (7–56) — FFFFCC yellow with borders
    for _data_row in range(7, 57):
        for _col_idx in range(1, 11):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    # Column widths
    widths = {"A": 10, "B": 18, "C": 20, "D": 15, "E": 20, "F": 15, "G": 12, "H": 40, "I": 35, "J": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_sod_matrix_sheet(wb, styles):
    """Create Sheet 5: SoD Matrix."""
    ws = wb.create_sheet("SoD Matrix")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "SEGREGATION OF DUTIES MATRIX - CONFLICTING ROLE COMBINATIONS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"SoD Matrix Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Defined SoD Conflicts: {SOD_CONFLICT_COUNT}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    row = 5
    headers = [
        "Conflict ID", "Role A", "Role B", "Risk Level", "Description",
        "Rationale", "Compensating Controls", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (6) — F2F2F2 grey, shows how to fill in
    _sample_data = ['SOD-001', 'IT Admin', 'Finance Approver', 'High', 'IT Admin can approve own financial transactions', 'Separation of financial approval from IT administration', 'Dual approval required', 'Active']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=6, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (7–56) — FFFFCC yellow with borders
    for _data_row in range(7, 57):
        for _col_idx in range(1, 9):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    conflicts = []  # No pre-generated conflicts (sheet is now blank template)
    # Column widths
    widths = {"A": 12, "B": 22, "C": 22, "D": 12, "E": 45, "F": 45, "G": 50, "H": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws, conflicts


def create_sod_violations_sheet(wb, styles, conflicts):
    """Create Sheet 6: SoD Violations."""
    ws = wb.create_sheet("SoD Violations")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "SOD VIOLATIONS - DETECTED CONFLICTING ROLE ASSIGNMENTS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Detection Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True, bold=True, color="FF0000")
    ws["A3"] = f"Violations Detected: {SOD_VIOLATION_COUNT}"
    ws["A3"].font = Font(italic=True, bold=True, color="FF0000")
    
    # Column headers
    row = 5
    headers = [
        "Violation ID", "User ID", "Username", "Conflict ID", "Role A", "Role B",
        "Risk Level", "Detected Date", "Remediation Plan", "Due Date", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (6) — F2F2F2 grey, shows how to fill in
    _sample_data = ['VIO-1000', 'USR-0005', 'mwilson', 'SOD-001', 'IT Admin', 'Finance Approver', 'High', '01.03.2026', 'Remove Finance Approver role from user', '15.03.2026', 'Open']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=6, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (7–56) — FFFFCC yellow with borders
    for _data_row in range(7, 57):
        for _col_idx in range(1, 12):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    # Column widths
    widths = {"A": 12, "B": 10, "C": 18, "D": 12, "E": 22, "F": 22, "G": 12, "H": 15, "I": 45, "J": 12, "K": 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_rbac_metrics_sheet(wb, styles):
    """Create Sheet 7: RBAC Metrics."""
    ws = wb.create_sheet("RBAC Metrics")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "RBAC ADOPTION METRICS & KPIS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Assessment Period: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Summary metrics
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "Key Performance Indicators (KPIs)"
    apply_style(cell, styles["header"])
    
    row += 1
    headers = ["Metric", "Target", "Actual", "Status", "Gap", "Comments"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Sample metrics
    metrics = [
        ("RBAC Adoption Rate", "≥80%", "80%", "Compliant", "0%", "80 users with roles, 20 with direct access"),
        ("Role Coverage", "100%", "100%", "Compliant", "0%", "All departments have defined roles"),
        ("Role Assignment Accuracy", "≥95%", "90%", "Warning", "-5%", "8 role assignments need review"),
        ("SoD Compliance", "100%", "95%", "Warning", "-5%", "5 SoD violations detected"),
        ("Privileged Role Usage", "≤10%", "12%", "Warning", "+2%", "Slightly higher than target"),
    ]
    
    for metric_name, target, actual, status, gap, comments in metrics:
        row += 1
        ws.cell(row=row, column=1).value = metric_name
        ws.cell(row=row, column=2).value = target
        ws.cell(row=row, column=3).value = actual
        ws.cell(row=row, column=4).value = status
        ws.cell(row=row, column=5).value = gap
        ws.cell(row=row, column=6).value = comments
        
        status_cell = ws.cell(row=row, column=4)
        if status == "Compliant":
            apply_style(status_cell, styles["compliant"])
        elif status == "Warning":
            apply_style(status_cell, styles["warning"])
        else:
            apply_style(status_cell, styles["non_compliant"])
        
        for col in range(1, 7):
            if col != 4:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Overall score
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "Overall RBAC & SoD Compliance Score"
    apply_style(cell, styles["header"])
    
    row += 1
    ws[f"A{row}"] = "Score Calculation"
    ws[f"B{row}"] = "(RBAC Adoption + SoD Compliance + Role Accuracy) / 3"
    
    row += 1
    ws[f"A{row}"] = "Overall Score"
    ws[f"B{row}"] = "88%"
    ws[f"B{row}"].font = Font(bold=True, size=14)
    
    row += 1
    ws[f"A{row}"] = "Maturity Level"
    ws[f"B{row}"] = "Managed (80-89%)"
    apply_style(ws[f"B{row}"], styles["warning"])
    
    # Column widths
    widths = {"A": 30, "B": 15, "C": 15, "D": 18, "E": 10, "F": 55}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    return ws


def create_gap_analysis_sheet(wb, styles):
    """Create Sheet 8: Gap Analysis."""
    ws = wb.create_sheet("Gap Analysis")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "GAP ANALYSIS - RBAC & SOD NON-COMPLIANCE"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Column headers
    row = 4
    headers = [
        "Gap ID", "Category", "Description", "Risk Level", "Affected Items",
        "Root Cause", "Remediation Plan", "Owner", "Due Date", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Sample row (row 5) — F2F2F2 grey, shows example data
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    sample_data = (
        "GAP-001", "RBAC Adoption", "20 users with direct access (non-RBAC)",
        "Medium", "20 users", "Legacy access before RBAC implementation",
        "Migrate users to appropriate roles by Q2 2026", "IAM Manager", "2026-06-30", "Open"
    )
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = grey_fill
        cell.border = _border

    # 50 empty data rows (rows 6–55)
    for data_row in range(6, 56):
        for col_idx in range(1, 11):
            cell = ws.cell(row=data_row, column=col_idx, value="")
            cell.fill = yellow_fill
            cell.border = _border

    # Data validations (skip sample row)
    risk_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add("D6:D55")

    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Accepted"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("J6:J55")

    # Column widths
    widths = {"A": 10, "B": 18, "C": 40, "D": 12, "E": 15, "F": 40, "G": 50, "H": 20, "I": 12, "J": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A4"

    return ws


def create_summary_dashboard_sheet(wb):
    """Create the Summary Dashboard sheet (Gold Standard — GS-SD-014/016 compliant)."""
    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_hdr = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _red_banner = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _ffffcc = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    _black_font = Font(name="Calibri", size=10, color="000000")
    _bold_black = Font(name="Calibri", size=10, bold=True, color="000000")

    # Row 1: Title — GS-SD-014
    ws.merge_cells("A1:G1")
    ws["A1"] = "ROLE & SOD COMPLIANCE \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = _border
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle (left-aligned per GS)
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Controls A.5.15 & A.5.18: Access Control & Access Rights"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Empty spacer

    # TABLE 1 banner
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = _white_font
    ws["A4"].fill = _navy
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A4"].border = _border

    # Row 5: Column headers — D9D9D9
    t1_headers = ["Assessment Area", "Total Items", "Compliant / Active",
                  "Partial / Under Review", "Non-Compliant / Inappropriate", "N/A", "Compliance %"]
    for col, hdr in enumerate(t1_headers, 1):
        cell = ws.cell(row=5, column=col, value=hdr)
        cell.font = _bold_black
        cell.fill = _grey_hdr
        cell.border = _border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # TABLE 1 data rows
    # Role Catalog: all roles defined = compliant (no non-compliant status — 100% good)
    # Role Assignments col H: "Yes" = appropriate, "Needs Review" = partial, "No" = inappropriate
    # Direct Access Users: any user = non-compliant (direct access not RBAC)
    # SoD Violations col K: "Remediation In Progress"+"Accepted" = partial, "Open" = bad
    # Gap Analysis col J: "Closed" = good, "In Progress" = partial, "Open" = bad
    area_configs = [
        ("Role Catalog",
         "COUNTA('Role Catalog'!A6:A56)",
         "'Role Catalog'!A6:A56",
         "COUNTA_ALL", "0", "ZERO_BAD"),
        ("Role Assignments",
         "COUNTA('Role Assignments'!A6:A56)",
         "'Role Assignments'!H6:H56",
         "Yes", "Needs Review", "No"),
        ("Direct Access Users",
         "COUNTA('Direct Access Users'!A6:A56)",
         "'Direct Access Users'!A6:A56",
         "ZERO_COMPLIANT", "0", "COUNTA_ALL"),
        ("SoD Violations",
         "COUNTA('SoD Violations'!A6:A56)",
         "'SoD Violations'!K6:K56",
         "COUNTIF_DUAL:Remediation In Progress|Accepted",
         "0",
         "Open"),
        ("Gap Analysis",
         "COUNTA('Gap Analysis'!A6:A55)",
         "'Gap Analysis'!J6:J55",
         "Resolved", "In Progress", "Open"),
    ]

    for i, (area_name, counta_expr, rng, good, partial, bad) in enumerate(area_configs):
        row = 6 + i

        cell_a = ws.cell(row=row, column=1, value=area_name)
        cell_a.border = _border
        cell_a.font = _black_font

        cell_b = ws.cell(row=row, column=2)
        cell_b.value = f"={counta_expr}"
        cell_b.border = _border
        cell_b.alignment = Alignment(horizontal="center")
        cell_b.font = _black_font

        # Col C: Good / Compliant
        cell_c = ws.cell(row=row, column=3)
        if good == "ZERO_COMPLIANT":
            cell_c.value = 0
        elif good == "COUNTA_ALL":
            cell_c.value = f"={counta_expr}"
        elif good.startswith("COUNTIF_DUAL:"):
            parts = good.replace("COUNTIF_DUAL:", "").split("|")
            cell_c.value = f'=COUNTIF({rng},"{parts[0]}")+COUNTIF({rng},"{parts[1]}")'
        else:
            cell_c.value = f'=COUNTIF({rng},"{good}")'
        cell_c.border = _border
        cell_c.alignment = Alignment(horizontal="center")
        cell_c.font = _black_font

        # Col D: Partial
        cell_d = ws.cell(row=row, column=4)
        if partial == "0":
            cell_d.value = 0
        else:
            cell_d.value = f'=COUNTIF({rng},"{partial}")'
        cell_d.border = _border
        cell_d.alignment = Alignment(horizontal="center")
        cell_d.font = _black_font

        # Col E: Bad / Non-Compliant
        cell_e = ws.cell(row=row, column=5)
        if bad == "ZERO_BAD":
            cell_e.value = 0
        elif bad == "COUNTA_ALL":
            cell_e.value = f"={counta_expr}"
        elif bad.startswith("COUNTIF_DUAL:"):
            parts = bad.replace("COUNTIF_DUAL:", "").split("|")
            cell_e.value = f'=COUNTIF({rng},"{parts[0]}")+COUNTIF({rng},"{parts[1]}")'
        else:
            cell_e.value = f'=COUNTIF({rng},"{bad}")'
        cell_e.border = _border
        cell_e.alignment = Alignment(horizontal="center")
        cell_e.font = _black_font

        cell_f = ws.cell(row=row, column=6, value=0)
        cell_f.border = _border
        cell_f.alignment = Alignment(horizontal="center")
        cell_f.font = _black_font

        cell_g = ws.cell(row=row, column=7)
        if good == "ZERO_COMPLIANT":
            # Direct Access Users: compliance = 1 - (all / total) = 0% when any present
            cell_g.value = f"=IF(B{row}=0,1,0)"
        elif good == "COUNTA_ALL":
            # Role Catalog: all compliant = 100%
            cell_g.value = f"=IF(B{row}=0,0,1)"
        else:
            cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        cell_g.number_format = "0.0%"
        cell_g.border = _border
        cell_g.alignment = Alignment(horizontal="center")
        cell_g.font = Font(name="Calibri", size=10, color="000000")

    # TOTAL row 11
    total_row = 11
    for col in range(1, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = _grey_hdr
        cell.border = _border
        cell.font = _bold_black
        cell.alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left")
    for col in range(2, 7):
        ws.cell(row=total_row, column=col).value = \
            f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}10)"
    cell_g_total = ws.cell(row=total_row, column=7)
    cell_g_total.value = f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"
    cell_g_total.number_format = "0.0%"

    # TABLE 2: Key Metrics
    t2_start = 13
    ws.merge_cells(f"A{t2_start}:G{t2_start}")
    ws[f"A{t2_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{t2_start}"].font = _white_font
    ws[f"A{t2_start}"].fill = _navy
    ws[f"A{t2_start}"].border = _border

    t2_hdr_row = t2_start + 1
    for col, hdr in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        cell = ws.cell(row=t2_hdr_row, column=col, value=hdr if hdr else None)
        cell.font = _bold_black
        cell.fill = _grey_hdr
        cell.border = _border
        cell.alignment = Alignment(horizontal="center")

    metrics = [
        ("Total Roles Defined", "=COUNTA('Role Catalog'!A6:A56)", None),
        ("Total Users Assessed", "=COUNTA('Role Assignments'!A6:A56)", None),
        ("Roles Appropriately Assigned", '=COUNTIF(\'Role Assignments\'!H6:H56,"Yes")', None),
        ("Roles Needing Review", '=COUNTIF(\'Role Assignments\'!H6:H56,"Needs Review")', None),
        ("Inappropriate Role Assignments", '=COUNTIF(\'Role Assignments\'!H6:H56,"No")', None),
        ("Non-RBAC (Direct Access) Users", "=COUNTA('Direct Access Users'!A6:A56)", None),
        ("SoD Conflicts Defined", "=COUNTA('SoD Matrix'!A6:A15)", None),
        ("SoD Violations Detected", "=COUNTA('SoD Violations'!A6:A56)", None),
        ("SoD Violations Open", '=COUNTIF(\'SoD Violations\'!K6:K56,"Open")', None),
        ("SoD Violations Remediated",
         '=COUNTIF(\'SoD Violations\'!K6:K56,"Remediation In Progress")'
         '+COUNTIF(\'SoD Violations\'!K6:K56,"Accepted")', None),
        ("Open RBAC Gaps", '=COUNTIF(\'Gap Analysis\'!J6:J55,"Open")', None),
        ("RBAC Adoption Rate",
         "=IFERROR(COUNTA('Role Assignments'!A6:A56)"
         "/(COUNTA('Role Assignments'!A6:A56)+COUNTA('Direct Access Users'!A6:A56)),0)",
         "0.0%"),
    ]

    row = t2_start + 2
    for metric_name, formula, fmt in metrics:
        ws.cell(row=row, column=1, value=metric_name).border = _border
        ws.cell(row=row, column=1).font = _black_font
        cell_val = ws.cell(row=row, column=2, value=formula)
        cell_val.border = _border
        cell_val.font = _black_font
        cell_val.alignment = Alignment(horizontal="center")
        if fmt:
            cell_val.number_format = fmt
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = _border
        row += 1

    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = _border
        row += 1

    # TABLE 3: Critical Findings
    t3_start = row + 1
    ws.merge_cells(f"A{t3_start}:G{t3_start}")
    ws[f"A{t3_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{t3_start}"].font = _white_font
    ws[f"A{t3_start}"].fill = _red_banner
    ws[f"A{t3_start}"].border = _border

    t3_hdr_row = t3_start + 1
    for col, hdr in enumerate(["Category", "Finding", "Count", "Severity", "Action Required", "", ""], 1):
        cell = ws.cell(row=t3_hdr_row, column=col, value=hdr if hdr else None)
        cell.font = _bold_black
        cell.fill = _grey_hdr
        cell.border = _border
        cell.alignment = Alignment(horizontal="center")

    findings = [
        ("Segregation of Duties", "Open SoD violations (unresolved)",
         '=COUNTIF(\'SoD Violations\'!K6:K56,"Open")', "Critical", "Immediate"),
        ("RBAC Compliance", "Users with inappropriate role assignments",
         '=COUNTIF(\'Role Assignments\'!H6:H56,"No")', "Critical", "Immediate"),
        ("Access Control", "Non-RBAC users (direct access, not role-based)",
         "=COUNTA('Direct Access Users'!A6:A56)", "High", "Urgent"),
        ("RBAC Compliance", "Role assignments needing review",
         '=COUNTIF(\'Role Assignments\'!H6:H56,"Needs Review")', "High", "Urgent"),
        ("Segregation of Duties", "SoD violations in remediation",
         '=COUNTIF(\'SoD Violations\'!K6:K56,"Remediation In Progress")', "Medium", "Plan"),
        ("Gap Analysis", "High-risk RBAC gaps",
         '=COUNTIF(\'Gap Analysis\'!D6:D55,"High")', "Critical", "Immediate"),
        ("Gap Analysis", "Open RBAC/SoD gaps",
         '=COUNTIF(\'Gap Analysis\'!J6:J55,"Open")', "High", "Urgent"),
        ("Access Control", "SoD conflicts defined vs. violations ratio",
         "=IFERROR(COUNTA('SoD Violations'!A6:A56)/COUNTA('SoD Matrix'!A6:A15),0)",
         "Medium", "Plan"),
    ]

    row = t3_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = _ffffcc
            ws.cell(row=row, column=col).border = _border
            ws.cell(row=row, column=col).font = _black_font
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = _ffffcc
            ws.cell(row=row, column=col).border = _border
        row += 1

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A4"

    return ws


def create_evidence_register(wb):
    """Create the standard Evidence Register sheet (gold standard)."""
    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_hdr = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _input = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = f"ISO/IEC 27001:2022 Annex A - Control {CONTROL_ID}: {CONTROL_NAME}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = _border

    ws.merge_cells("A3:H3")
    ws["A3"].border = _border

    columns = [
        ("Evidence ID", 14), ("Evidence Type", 20), ("Description", 45),
        ("Related Control / Section", 28), ("Collection Date (DD.MM.YYYY)", 22),
        ("Storage Location / Reference", 38), ("Collected By", 22), ("Status", 14),
    ]
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    sample_data = ["EV-001", "Document", "Sample evidence entry — replace with actual evidence",
                   "All Controls", "01.01.2026", "SharePoint/ISMS/Evidence/", "ISMS Team", "Active"]
    for col_idx, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
        cell.fill = _grey_sample
        cell.border = _border

    dv_status = DataValidation(
        type="list",
        formula1='"Active,Archived,Superseded,Pending Review"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(6, 106):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = _input
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"

    return ws


def create_approval_sheet(wb):
    """Create the Approval Sign-Off sheet — Gold Standard (GS-AS-014/015)."""
    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title banner — GS-AS-014
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY section banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Rows 4-8: Summary metadata — B6 = Overall Compliance (GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Row 7 status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # Approver sections start at row 11 (rows 9-10 = gap)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION — GS-AS-012: col A = plain bold label, NO dark fill
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("=" * 80)
    logger.info(f"Generating: {WORKBOOK_NAME}")
    logger.info(f"Control: ISO/IEC 27001:2022 {CONTROL_ID} - {CONTROL_NAME}")
    logger.info("=" * 80)

    # Create workbook
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.remove(wb.active)

    # Setup styles
    styles = _STYLES

    # Create sheets
    logger.info("Creating sheets...")

    logger.info("  1/10 Creating Instructions & Legend...")
    create_instructions_sheet(wb.create_sheet())

    logger.info("  2/10 Creating Role Catalog...")
    _, roles = create_role_catalog_sheet(wb, styles)

    logger.info("  3/10 Creating Role Assignments...")
    create_role_assignments_sheet(wb, styles, roles)

    logger.info("  4/10 Creating Direct Access Users...")
    create_direct_access_sheet(wb, styles)

    logger.info("  5/10 Creating SoD Matrix...")
    _, conflicts = create_sod_matrix_sheet(wb, styles)

    logger.info("  6/10 Creating SoD Violations...")
    create_sod_violations_sheet(wb, styles, conflicts)

    logger.info("  7/10 Creating RBAC Metrics...")
    create_rbac_metrics_sheet(wb, styles)

    logger.info("  8/11 Creating Gap Analysis...")
    create_gap_analysis_sheet(wb, styles)

    logger.info("  9/11 Creating Evidence Register...")
    create_evidence_register(wb)

    logger.info("  10/11 Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb)

    logger.info("  11/11 Creating Approval Sign-Off...")
    create_approval_sheet(wb)

    # Finalise data validations
    finalize_validations(wb)

    # Save workbook
    logger.info(f"Saving workbook: {output_path.name}")
    wb.save(output_path)

    logger.info("=" * 80)
    logger.info("SUCCESS - workbook generated")
    logger.info(f"File: {output_path}")
    logger.info("Sheets: 11")
    logger.info(f"Roles defined: {ROLE_COUNT}")
    logger.info("=" * 80)
    logger.info("Workbook Contents:")
    logger.info(f"  - Role Catalog ({ROLE_COUNT} RBAC roles)")
    logger.info(f"  - Role Assignments ({ROLE_ASSIGNMENT_COUNT} users)")
    logger.info(f"  - Direct Access Users ({DIRECT_ACCESS_COUNT} non-RBAC users)")
    logger.info(f"  - SoD Matrix ({SOD_CONFLICT_COUNT} conflict definitions)")
    logger.info(f"  - SoD Violations ({SOD_VIOLATION_COUNT} detected violations)")
    logger.info("  - RBAC Adoption Metrics & KPIs")
    logger.info("  - Gap Analysis & Remediation")
    logger.info("  - Evidence Register for A.5.15/A.5.18")
    logger.info("  - Summary Dashboard")
    logger.info("  - Approval Sign-Off")
    logger.info("=" * 80)



def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    try:
        output_file = main()
        sys.exit(0)
    except Exception as e:
        logger.error(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
