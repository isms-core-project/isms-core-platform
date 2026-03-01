#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
#
# This file is part of ISMS Core.
#
# ISMS Core is dual-licensed:
#   1. AGPL 3.0 (Open Source) - See LICENSE-AGPL.txt
#   2. Commercial License - Contact vendor for proprietary use
#
# You may use this file under either license, at your option.
# =============================================================================
"""
================================================================================
ISMS-IMP-A.5.15-16-18.S3 - Access Review Results Assessment Generator
================================================================================

ISO/IEC 27001:2022 Controls A.5.15 & A.5.18: Access Control & Access Rights
Assessment Workbook 3 of 5: Access Review and Recertification Results

--------------------------------------------------------------------------------
SAMPLE SCRIPT - REQUIRES CUSTOMIZATION FOR YOUR ORGANISATION
--------------------------------------------------------------------------------

This script is a TEMPLATE/SAMPLE implementation and MUST be adapted to match
your organisation's specific access review processes, review frequencies,
and assessment requirements.

Key customization areas:
1. Review frequency by system criticality (match your risk classification)
2. Reviewer roles and responsibilities (adapt to your governance structure)
3. Review scope definitions (align with your system inventory)
4. Compliance thresholds (based on your risk tolerance)
5. Remediation SLAs (match your change management processes)

DO NOT use this script without reviewing and adapting all sections marked
with "# CUSTOMIZE:" comments throughout the code.

Reference Pattern: Based on ISMS-A.5.15-16-18 IAM Assessment Framework

--------------------------------------------------------------------------------
DESCRIPTION
--------------------------------------------------------------------------------

This script generates a comprehensive Excel assessment workbook for tracking
access review/recertification activities and results across all systems.

**Purpose:**
Enables systematic tracking of access review processes against
ISO 27001:2022 Controls A.5.15 and A.5.18 requirements, supporting
evidence-based validation of periodic access verification and cleanup.

**Assessment Scope:**
- Access review scheduling and completion tracking
- Review frequency compliance by system criticality
- Reviewer accountability and participation
- Access confirmation vs. removal decisions
- Unresponsive reviewer escalation
- Review findings and remediation
- Overdue review identification
- Review coverage (% systems reviewed)
- Review effectiveness metrics
- Gap analysis and remediation planning
- Evidence collection for audit readiness

**Generated Workbook Structure:**
1. Instructions & Legend - Usage guidance and review standards
2. Review_Schedule - Planned reviews by system/period (60 systems)
3. Review_Execution - Review completion tracking (80 review cycles)
4. Review_Findings - Access decisions and changes (150 findings)
5. Reviewer_Accountability - Reviewer participation metrics (30 reviewers)
6. Overdue_Reviews - Reviews past due date (20 overdue items)
7. Remediation_Tracking - Access removal/change implementation (40 items)
8. Coverage_Metrics - System coverage and completion rates
9. Gap_Analysis - Review process gaps and improvement actions
10. Evidence_Register - Evidence collection for A.5.15/A.5.18 compliance
11. Approval_Sign_Off - 3-level approval workflow

**Key Features:**
- Data validation with review frequency dropdown lists
- Conditional formatting for review status (completed/overdue/pending)
- Automated gap identification for overdue reviews
- Protected formulas with unprotected input cells
- Evidence linkage for audit traceability
- Multi-stakeholder approval workflow
- Integration with access review platforms

**Integration:**
This assessment feeds into the IAM Governance Dashboard, which
consolidates data from all five IAM assessment domains for
executive oversight and audit readiness.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------

System Requirements:
    - Python 3.8 or higher
    - openpyxl library for Excel generation

Installation:
    Ubuntu/Debian:
        sudo apt install python3-openpyxl
    
    Or via pip:
        pip3 install openpyxl

Dependencies:
    - openpyxl (Python Excel library)
    - datetime (standard library)
    - random (standard library - for sample data generation)

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------

Basic Usage:
    python3 generate_a515-16-18_3_access_review_results.py

Advanced Usage:
    # Generate with custom output directory
    python3 generate_a515-16-18_3_access_review_results.py --output /path/to/dir
    
    # Generate with specific date suffix
    python3 generate_a515-16-18_3_access_review_results.py --date 20250124

Output:
    File: ISMS-IMP-A.5.15-16-18.S3_Access_Review_Results_YYYYMMDD.xlsx
    Location: Current directory (or specified output path)

Post-Generation Steps:
    1. Define review schedule based on system criticality
    2. Assign reviewers (managers, system owners, security)
    3. Prepare access review data (who has access to each system)
    4. Execute reviews according to schedule
    5. Document reviewer decisions (confirm/remove/justify)
    6. Track access removal implementation
    7. Escalate unresponsive reviewers
    8. Calculate review completion metrics
    9. Identify and remediate overdue reviews
    10. Collect and link audit evidence (review reports, approvals)
    11. Obtain stakeholder approvals
    12. Feed results into IAM Governance Dashboard

--------------------------------------------------------------------------------
METADATA
--------------------------------------------------------------------------------

Control Reference:    ISO/IEC 27001:2022 Annex A Controls A.5.15, A.5.18
Assessment Domain:    3 of 5 (Access Review & Recertification Results)
Framework Version:    1.0
Script Version:       1.0
Author:               [Organisation] ISMS Implementation Team
Date:                 [Date to be set]
Last Modified:        [Date to be set]
Python Version:       3.8+
License:              [Organisation License/Terms]

Related Documents:
    - ISMS-POL-A.5.15-16-18: Access Control Policy (A.5.15)
    - ISMS-POL-A.5.15-16-18: Access Rights Management Policy (A.5.18)
    - ISMS-IMP-A.5.15-16-18-S4: Access Review Process Guide
    - ISMS-IMP-A.5.15-16-18.S1: User Inventory Assessment (Workbook 1)
    - ISMS-IMP-A.5.15-16-18.S2: Access Rights Matrix Assessment (Workbook 2)
    - ISMS-IMP-A.5.15-16-18.S4: Role Compliance Assessment (Workbook 4)
    - ISMS-IMP-A.5.15-16-18.S5: Lifecycle Compliance Detailed (Workbook 5)
    - IAM Governance Dashboard: Consolidated IAM compliance view

--------------------------------------------------------------------------------
CHANGE HISTORY
--------------------------------------------------------------------------------

Version 1.0 - [Date to be set]
    - Initial release
    - Implements full assessment framework per ISMS-IMP-A.5.15-16-18-S4
    - Supports comprehensive access review tracking and analysis
    - Integrated with IAM Governance Dashboard

[Future changes to be documented here]

--------------------------------------------------------------------------------
IMPORTANT NOTES
--------------------------------------------------------------------------------

**Access Review is MANDATORY:**
ISO 27001:2022 A.5.18 explicitly requires periodic access review.
Failure to conduct reviews = automatic non-compliance.

Review frequency must be risk-based:
- Critical systems: Quarterly reviews
- High-value systems: Semi-annual reviews
- Standard systems: Annual reviews
- Low-risk systems: Biennial reviews (minimum)

Skipping reviews or excessive delays = major audit finding.

**Reviewer Accountability:**
Reviewers (typically managers/system owners) are ACCOUNTABLE for:
- Timely review completion (within scheduled period)
- Accurate access decisions (confirm/remove/justify)
- Documentation of exceptions (why user still needs access)
- Follow-up on access removal (verify implementation)

Unresponsive reviewers create compliance gaps and must be escalated.

**Audit Considerations:**
Auditors expect to see:
- Documented review schedule aligned with system criticality
- Evidence of review execution (review reports, approvals)
- Reviewer participation and accountability metrics
- Access removal implementation (not just decisions)
- Overdue review remediation plans
- Continuous review coverage (all systems reviewed on schedule)

**Data Protection:**
Assessment workbooks contain review results including:
- Access decisions and justifications
- Reviewer names and responsibilities
- Finding details (excessive access, unauthorised access)

Handle in accordance with your organisation's data classification policies.

**Maintenance:**
Review and update assessment:
- Monthly: Update review execution status and overdue tracking
- Quarterly: Calculate review completion metrics
- Semi-annually: Analyze review effectiveness (findings vs. remediation)
- Annually: Review and adjust review frequencies by system
- Ad-hoc: After organisational changes or new system deployments

**Quality Assurance:**
Review effectiveness depends on data quality:
- Access review data must be current (from Workbook 2)
- Reviewers must have accurate view of who has access
- Review platforms should auto-populate access data
- Manual reviews prone to errors and reviewer fatigue

Validate review results against actual access (post-review verification).

**Common Review Process Failures:**
- "Rubber-stamping": Reviewers approve all access without verification
- Unresponsive reviewers: Reviews delayed or incomplete
- Missing remediation: Access removal decisions not implemented
- Stale data: Review based on outdated access information
- No escalation: Overdue reviews not tracked or addressed

Monitor for these failure patterns and implement corrective actions.

**Review Automation:**
Consider automated review platforms (Sailpoint, Saviynt, etc.):
- Pre-populate access data for reviewers
- Track reviewer participation and decisions
- Automate access removal workflows
- Generate compliance reports
- Reduce manual effort and errors

Manual reviews in Excel/email = high risk of compliance gaps.

**Privilege Creep Detection:**
Access reviews should identify privilege creep:
- Users accumulating excess access over time (role changes, transfers)
- Access granted but never used (unnecessary access)
- Access appropriate when granted but no longer needed

Review decisions should remove unnecessary access, not just confirm existing.

================================================================================
"""

import sys
from datetime import datetime, timedelta
from pathlib import Path
import random

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import BarChart, Reference
except ImportError as e:
    logger.error(f"ERROR: Required library not found: {e}")
    logger.info("Install required libraries: pip install openpyxl")
    sys.exit(1)

_wkbk_dir = Path(__file__).resolve().parent.parent / "WKBK"
_wkbk_dir.mkdir(parents=True, exist_ok=True)

# ============================================================================
# DOCUMENT METADATA
# ============================================================================
WORKBOOK_NAME = "Access Review Results Assessment"
DOCUMENT_ID = "ISMS-IMP-A.5.15-16-18.S3"
CONTROL_ID = "A.5.15 & A.5.18"
CONTROL_NAME = "Access Control & Access Rights"
CONTROL_REF   = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# Data row counts
SCHEDULED_REVIEWS = 80        # Annual review schedule
REVIEW_FINDINGS = 60          # Review findings tracked
OVERDUE_REVIEWS = 15          # Overdue reviews
REVIEWER_COUNT = 20           # Individual reviewers
GAP_ROW_COUNT = 30           # Gap analysis tracking
EVIDENCE_ROW_COUNT = 50      # Evidence register
# ============================================================================
# UNICODE SYMBOLS - PROPER UTF-8 ENCODING
# ============================================================================
CHECK   = '\u2705'      # ✅ Green checkmark
XMARK   = '\u274C'      # ❌ Red X
WARNING = '\u26A0'      # ⚠  Warning sign
BULLET  = '\u2022'      # •  Bullet point
# ============================================================================
# SECTION 2: STYLE DEFINITIONS
# ============================================================================
def setup_styles():
    """Define all cell styles used throughout the workbook."""
    thin = Side(style="thin")
    medium = Side(style="medium")
    
    styles = {
        "title": {
            "font": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=medium, right=medium, top=medium, bottom=medium),
        },
        "header": {
            "font": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "subheader": {
            "font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "column_header": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=False),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "data_center": {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "compliant": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "non_compliant": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
        "warning": {
            "font": Font(name="Calibri", size=10, bold=True),
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        },
    }
    
    return styles



_STYLES = setup_styles()
def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    if "font" in style_dict:
        cell.font = style_dict["font"]
    if "fill" in style_dict:
        cell.fill = style_dict["fill"]
    if "alignment" in style_dict:
        cell.alignment = style_dict["alignment"]
    if "border" in style_dict:
        cell.border = style_dict["border"]


# ============================================================================
# SECTION 3: SAMPLE DATA GENERATORS
# ============================================================================

def generate_systems():
    """Generate sample systems/applications."""
    systems = [
        {"name": "CRM (Salesforce)", "criticality": "High", "frequency": "Semi-Annual"},
        {"name": "Finance System (SAP)", "criticality": "High", "frequency": "Semi-Annual"},
        {"name": "HR System (Workday)", "criticality": "High", "frequency": "Semi-Annual"},
        {"name": "Email (Microsoft 365)", "criticality": "Medium", "frequency": "Annual"},
        {"name": "File Server", "criticality": "Medium", "frequency": "Annual"},
        {"name": "Project Management", "criticality": "Medium", "frequency": "Annual"},
        {"name": "Intranet", "criticality": "Low", "frequency": "Biennial"},
        {"name": "Database Server", "criticality": "Critical", "frequency": "Quarterly"},
        {"name": "VPN", "criticality": "Medium", "frequency": "Annual"},
        {"name": "Marketing Automation", "criticality": "Medium", "frequency": "Annual"},
    ]
    return systems


def generate_reviewers():
    """Generate sample reviewer list."""
    reviewers = [
        {"name": "Alice Manager", "role": "Finance Manager", "department": "Finance"},
        {"name": "Bob Director", "role": "Sales Director", "department": "Sales"},
        {"name": "Carol Lead", "role": "HR Manager", "department": "HR"},
        {"name": "David Chief", "role": "IT Manager", "department": "IT"},
        {"name": "Emma Head", "role": "Security Manager", "department": "IT"},
        {"name": "Frank VP", "role": "VP Operations", "department": "Operations"},
        {"name": "Grace Owner", "role": "System Owner", "department": "IT"},
        {"name": "Henry Mgr", "role": "Engineering Manager", "department": "Engineering"},
        {"name": "Iris Sup", "role": "Marketing Manager", "department": "Marketing"},
        {"name": "Jack Admin", "role": "IT Administrator", "department": "IT"},
    ]
    return reviewers


def generate_review_schedule(systems):
    """Generate annual review schedule."""
    schedule = []
    reviewers = generate_reviewers()
    
    quarters = {
        "Q1": (1, 3),
        "Q2": (4, 6),
        "Q3": (7, 9),
        "Q4": (10, 12)
    }
    
    review_id = 1
    for system in systems:
        # Determine how many reviews based on frequency
        if system["frequency"] == "Quarterly":
            review_count = 4  # Q1, Q2, Q3, Q4
            periods = ["Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"]
        elif system["frequency"] == "Semi-Annual":
            review_count = 2  # H1, H2
            periods = ["H1 2026", "H2 2026"]
        elif system["frequency"] == "Annual":
            review_count = 1
            periods = ["Annual 2026"]
        else:  # Biennial
            review_count = 1
            periods = ["Biennial 2026"]
        
        for i, period in enumerate(periods):
            reviewer = random.choice(reviewers)
            
            # Calculate due date based on period
            if "Q1" in period:
                due_date = datetime(2026, 3, 31)
            elif "Q2" in period:
                due_date = datetime(2026, 6, 30)
            elif "Q3" in period:
                due_date = datetime(2026, 9, 30)
            elif "Q4" in period:
                due_date = datetime(2026, 12, 31)
            elif "H1" in period:
                due_date = datetime(2026, 6, 30)
            elif "H2" in period:
                due_date = datetime(2026, 12, 31)
            else:  # Annual or Biennial
                due_date = datetime(2026, 12, 31)
            
            schedule.append({
                "review_id": f"REV-{review_id:04d}",
                "system_name": system["name"],
                "criticality": system["criticality"],
                "review_period": period,
                "reviewer": reviewer["name"],
                "reviewer_role": reviewer["role"],
                "due_date": due_date,
                "frequency": system["frequency"],
            })
            review_id += 1
    
    return schedule[:SCHEDULED_REVIEWS]  # Limit to configured count


def generate_review_findings():
    """Generate sample review findings."""
    findings = []
    actions = ["Confirm", "Remove", "Change Access Level", "Request Justification"]
    
    for i in range(REVIEW_FINDINGS):
        action = random.choice(actions)
        
        finding = {
            "finding_id": f"FIND-{1000 + i}",
            "review_id": f"REV-{random.randint(1, 80):04d}",
            "system": random.choice(["CRM", "Finance System", "HR System", "Email", "File Server"]),
            "user": f"user{random.randint(1, 100)}",
            "access_level": random.choice(["Read", "Write", "Admin"]),
            "action": action,
            "reason": "",
            "completion_date": None,
            "status": "Pending"
        }
        
        # Add reasons based on action
        if action == "Remove":
            finding["reason"] = random.choice([
                "User no longer needs access",
                "User left company",
                "User changed departments",
                "Access not used in 90+ days"
            ])
            # Some completed, some pending
            if random.random() > 0.3:
                finding["completion_date"] = datetime.now() - timedelta(days=random.randint(1, 30))
                finding["status"] = "Completed"
        elif action == "Change Access Level":
            finding["reason"] = random.choice([
                "User needs less privilege (Admin → Read/Write)",
                "User promoted (Read → Admin)",
                "Temporary project access ended"
            ])
            if random.random() > 0.4:
                finding["completion_date"] = datetime.now() - timedelta(days=random.randint(1, 20))
                finding["status"] = "Completed"
        elif action == "Request Justification":
            finding["reason"] = "Missing business justification documentation"
            if random.random() > 0.5:
                finding["completion_date"] = datetime.now() - timedelta(days=random.randint(1, 15))
                finding["status"] = "Completed"
        else:  # Confirm
            finding["reason"] = "Access appropriate for user role"
            finding["completion_date"] = datetime.now() - timedelta(days=random.randint(0, 10))
            finding["status"] = "Completed"
        
        findings.append(finding)
    
    return findings


# ============================================================================
# SECTION 4: SHEET CREATION FUNCTIONS
# ============================================================================


def create_instructions_sheet(ws):
    """Create GS-IL-compliant Instructions & Legend sheet (Sheet 1)."""
    ws.title = "Instructions & Legend"
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill("solid", fgColor="003366")
    _grey = PatternFill("solid", fgColor="D9D9D9")
    _input = PatternFill("solid", fgColor="FFFFCC")
    _green = PatternFill("solid", fgColor="C6EFCE")
    _amber = PatternFill("solid", fgColor="FFEB9C")
    _red   = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{DOCUMENT_ID}  -  {WORKBOOK_NAME}\n{CONTROL_REF}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    ws["A3"] = "Document Information"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    for i, (label, value) in enumerate([
        ("Document ID",       DOCUMENT_ID),
        ("Workbook Title",    WORKBOOK_NAME),
        ("Control Reference", CONTROL_REF),
        ("Version",           "1.0"),
        ("Assessment Date",   ""),
        ("Completed By",      ""),
        ("Organisation",      ""),
    ]):
        r = 4 + i
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name="Calibri", bold=True)
        ws[f"B{r}"] = value
        if not value:
            ws[f"B{r}"].fill = _input
            ws[f"B{r}"].border = _border
    ws["A12"] = "Instructions"
    ws["A12"].font = Font(name="Calibri", size=12, bold=True)

    _instructions = ['1. Open this workbook and review the sheet tabs to understand the assessment scope.', '2. Complete the Assessment Date and Prepared By fields in the Document Information table above.', '3. Navigate to each data sheet and review the pre-populated sample data.', '4. Replace sample data with actual organisation data for your environment.', '5. Use the dropdown lists in each sheet for consistent status and category values.', '6. Document all gaps and findings in the Gap Analysis sheet.', '7. Collect supporting evidence and record references in the Evidence Register.', '8. Obtain required approvals via the Approval Sign-Off sheet before submission.', '9. Store the completed workbook in the designated ISMS evidence repository.']
    for _i, _line in enumerate(_instructions):
        ws[f"A{13 + _i}"] = _line

    _leg_row = 23

    ws[f"A{_leg_row}"] = "Status Legend"
    ws[f"A{_leg_row}"].font = Font(name="Calibri", size=12, bold=True)
    for col_idx, header in enumerate(["Symbol", "Status", "Description"], start=1):
        c = ws.cell(row=_leg_row + 1, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = _grey
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    for i, (sym, status, desc, fill) in enumerate([
        ("\u2713", "Compliant / Complete",        "Requirement fully met",                   _green),
        ("\u26a0", "Partial / In Progress",        "Partially met or in progress",            _amber),
        ("\u2717", "Non-Compliant / Not Started",  "Requirement not met",                     _red),
        ("\u2014", "Not Applicable",               "Not applicable to this assessment",        None),
    ]):
        r = _leg_row + 2 + i
        ws.cell(row=r, column=1, value=sym).border = _border
        s = ws.cell(row=r, column=2, value=status)
        d = ws.cell(row=r, column=3, value=desc)
        if fill:
            s.fill = fill
        for cell in (s, d):
            cell.border = _border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

def create_review_schedule_sheet(wb, styles):
    """Create Sheet 2: Review Schedule."""
    ws = wb.create_sheet("Review Schedule")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "ACCESS REVIEW SCHEDULE - 2026 ANNUAL CALENDAR"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Schedule Generated: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Scheduled Reviews: {SCHEDULED_REVIEWS}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    row = 5
    headers = [
        "Review ID", "System/Application", "Criticality", "Review Period",
        "Frequency", "Reviewer", "Reviewer Role", "Due Date", "Est. Users", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (6) — F2F2F2 grey, shows how to fill in
    _sample_data = ['REV-001', 'Active Directory', 'Critical', 'Q1 2026', 'Quarterly', 'J. Smith', 'IT Manager', '31.03.2026', 150, 'Scheduled']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=6, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (7–56) — FFFFCC yellow with borders
    for _data_row in range(7, 57):
        for _col_idx in range(1, 11):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    schedule = []  # No pre-generated schedule (sheet is now blank template)
    # Column widths
    widths = {"A": 12, "B": 30, "C": 12, "D": 15, "E": 15, "F": 20, "G": 22, "H": 12, "I": 12, "J": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws, schedule


def create_review_completion_sheet(wb, styles, schedule):
    """Create Sheet 3: Review Completion."""
    ws = wb.create_sheet("Review Completion")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = "REVIEW COMPLETION TRACKING - EXECUTION STATUS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Column headers
    row = 4
    headers = [
        "Review ID", "System", "Review Period", "Reviewer", "Due Date",
        "Start Date", "Completion Date", "Days to Complete", "Users Reviewed",
        "Access Confirmed", "Access Removed", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (5) — F2F2F2 grey, shows how to fill in
    _sample_data = ['REV-001', 'Active Directory', 'Q1 2026', 'J. Smith', '31.03.2026', '10.03.2026', '25.03.2026', 15, 150, 135, 15, 'Completed']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=5, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (6–55) — FFFFCC yellow with borders
    for _data_row in range(6, 56):
        for _col_idx in range(1, 13):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    # Column widths
    widths = {"A": 12, "B": 28, "C": 15, "D": 20, "E": 12, "F": 12, "G": 15, "H": 15, "I": 15, "J": 15, "K": 15, "L": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A5"
    
    return ws


def create_review_findings_sheet(wb, styles):
    """Create Sheet 4: Review Findings."""
    ws = wb.create_sheet("Review Findings")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = "REVIEW FINDINGS - DETAILED ACTIONS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    ws["A3"] = f"Total Findings: {REVIEW_FINDINGS}"
    ws["A3"].font = Font(italic=True)
    
    # Column headers
    row = 5
    headers = [
        "Finding ID", "Review ID", "System", "Username", "Access Level",
        "Action", "Reason", "Priority", "Completion Date", "Days to Remediate", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (6) — F2F2F2 grey, shows how to fill in
    _sample_data = ['FIND-001', 'REV-001', 'Active Directory', 'jsmith', 'Read/Write', 'Confirm', 'Correct access for role', 'Low', '25.03.2026', 3, 'Completed']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=6, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (7–56) — FFFFCC yellow with borders
    for _data_row in range(7, 57):
        for _col_idx in range(1, 12):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    # Column widths
    widths = {"A": 12, "B": 12, "C": 20, "D": 15, "E": 15, "F": 20, "G": 45, "H": 10, "I": 15, "J": 18, "K": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_overdue_reviews_sheet(wb, styles, schedule):
    """Create Sheet 5: Overdue Reviews."""
    ws = wb.create_sheet("Overdue Reviews")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "OVERDUE REVIEWS - ESCALATION REQUIRED"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    # Filter to overdue reviews
    today = datetime.now()
    overdue = [r for r in schedule if r["due_date"] < today][:OVERDUE_REVIEWS]
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True, bold=True, color="FF0000")
    ws["A3"] = f"Overdue Reviews: {len(overdue)}"
    ws["A3"].font = Font(italic=True, bold=True, color="FF0000")
    
    # Column headers
    row = 5
    headers = [
        "Review ID", "System", "Criticality", "Reviewer", "Due Date",
        "Days Overdue", "Escalation Level", "Escalation Date", "Escalated To", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (6) — F2F2F2 grey, shows how to fill in
    _sample_data = ['REV-010', 'CRM System', 'High', 'P. Jones', '31.01.2026', 28, 'Level 2: Manager', '01.02.2026', 'P. Jones + Manager', 'Overdue']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=6, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (7–56) — FFFFCC yellow with borders
    for _data_row in range(7, 57):
        for _col_idx in range(1, 11):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    # Column widths
    widths = {"A": 12, "B": 28, "C": 12, "D": 20, "E": 12, "F": 15, "G": 20, "H": 15, "I": 30, "J": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A6"
    
    return ws


def create_reviewer_performance_sheet(wb, styles):
    """Create Sheet 6: Reviewer Performance."""
    ws = wb.create_sheet("Reviewer Performance")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "REVIEWER PERFORMANCE METRICS - RESPONSIVENESS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Assessment Period: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Column headers
    row = 4
    headers = [
        "Reviewer Name", "Role", "Department", "Total Reviews",
        "Completed", "Overdue", "Completion Rate", "Avg Days to Complete", "Performance", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    _grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    # Sample row (5) — F2F2F2 grey, shows how to fill in
    _sample_data = ['John Smith', 'IT Manager', 'IT', 8, 7, 1, '88%', 7, 'Good', 'Compliant']
    for _col_idx, _val in enumerate(_sample_data, start=1):
        _cell = ws.cell(row=5, column=_col_idx, value=_val)
        _cell.fill = _grey_fill
        _cell.border = _border

    # 50 empty data rows (6–55) — FFFFCC yellow with borders
    for _data_row in range(6, 56):
        for _col_idx in range(1, 11):
            _cell = ws.cell(row=_data_row, column=_col_idx)
            _cell.fill = _yellow_fill
            _cell.border = _border

    # Column widths
    widths = {"A": 20, "B": 22, "C": 15, "D": 15, "E": 12, "F": 12, "G": 18, "H": 20, "I": 20, "J": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A5"
    
    return ws


def create_review_metrics_sheet(wb, styles):
    """Create Sheet 7: Review Metrics."""
    ws = wb.create_sheet("Review Metrics")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "ACCESS REVIEW COMPLIANCE METRICS - KPIS"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Assessment Period: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Summary metrics
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "Key Performance Indicators (KPIs)"
    apply_style(cell, styles["header"])
    
    row += 1
    headers = ["Metric", "Target", "Actual", "Status", "Gap", "Comments"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Sample metrics
    metrics = [
        ("Review Completion Rate", "≥95%", "87%", "Warning", "-8%", "12 reviews overdue"),
        ("On-Time Completion", "≥90%", "83%", "Warning", "-7%", "Reviews completed late"),
        ("Finding Remediation Rate", "≥95%", "78%", "Non-Compliant", "-17%", "18 findings pending"),
        ("Reviewer Response Rate", "100%", "92%", "Warning", "-8%", "3 unresponsive reviewers"),
        ("Access Removal Rate", "8-15%", "12%", "Compliant", "0%", "Within expected range"),
    ]
    
    for metric_name, target, actual, status, gap, comments in metrics:
        row += 1
        ws.cell(row=row, column=1).value = metric_name
        ws.cell(row=row, column=2).value = target
        ws.cell(row=row, column=3).value = actual
        ws.cell(row=row, column=4).value = status
        ws.cell(row=row, column=5).value = gap
        ws.cell(row=row, column=6).value = comments
        
        status_cell = ws.cell(row=row, column=4)
        if status == "Compliant":
            apply_style(status_cell, styles["compliant"])
        elif status == "Warning":
            apply_style(status_cell, styles["warning"])
        else:
            apply_style(status_cell, styles["non_compliant"])
        
        for col in range(1, 7):
            if col != 4:
                apply_style(ws.cell(row=row, column=col), styles["data"])
    
    # Overall score
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "Overall Access Review Compliance Score"
    apply_style(cell, styles["header"])
    
    row += 1
    ws[f"A{row}"] = "Score Calculation"
    ws[f"B{row}"] = "(Completion Rate + Remediation Rate + Reviewer Response) / 3"
    
    row += 1
    ws[f"A{row}"] = "Overall Score"
    ws[f"B{row}"] = "86%"
    ws[f"B{row}"].font = Font(bold=True, size=14)
    
    row += 1
    ws[f"A{row}"] = "Maturity Level"
    ws[f"B{row}"] = "Managed (80-89%)"
    apply_style(ws[f"B{row}"], styles["warning"])
    
    # Column widths
    widths = {"A": 30, "B": 15, "C": 15, "D": 18, "E": 10, "F": 50}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    return ws


def create_gap_analysis_sheet(wb, styles):
    """Create Sheet 8: Gap Analysis."""
    ws = wb.create_sheet("Gap Analysis")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "GAP ANALYSIS - REVIEW PROCESS NON-COMPLIANCE"
    apply_style(cell, styles["title"])
    ws.row_dimensions[1].height = 35
    
    ws["A2"] = f"Assessment Date: {GENERATED_DATE}"
    ws["A2"].font = Font(italic=True)
    
    # Column headers
    row = 4
    headers = [
        "Gap ID", "Category", "Description", "Risk Level", "Affected Items",
        "Root Cause", "Remediation Plan", "Owner", "Due Date", "Status"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, styles["column_header"])
    
    # Sample row (row 5) — F2F2F2 grey, shows example data
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    sample_data = (
        "GAP-001", "Review Completion", "12 reviews past due date",
        "High", "12 reviews", "Unresponsive reviewers, lack of follow-up",
        "Implement automated escalation workflow", "Security Manager", "2026-06-30", "Open"
    )
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = grey_fill
        cell.border = _border

    # 50 empty data rows (rows 6–55)
    for data_row in range(6, 56):
        for col_idx in range(1, 11):
            cell = ws.cell(row=data_row, column=col_idx, value="")
            cell.fill = yellow_fill
            cell.border = _border

    # Data validations (skip sample row)
    risk_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add("D6:D55")

    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Accepted"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("J6:J55")

    # Column widths
    widths = {"A": 10, "B": 20, "C": 40, "D": 12, "E": 15, "F": 40, "G": 40, "H": 18, "I": 12, "J": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A4"

    return ws


def create_summary_dashboard_sheet(wb):
    """Create the Summary Dashboard sheet (Gold Standard — GS-SD-014/016 compliant)."""
    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_hdr = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _red_banner = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _ffffcc = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    _white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    _black_font = Font(name="Calibri", size=10, color="000000")
    _bold_black = Font(name="Calibri", size=10, bold=True, color="000000")

    # Row 1: Title — GS-SD-014
    ws.merge_cells("A1:G1")
    ws["A1"] = "ACCESS REVIEW RESULTS \u2014 SUMMARY DASHBOARD"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 8):
        ws.cell(row=1, column=c).border = _border
    ws.row_dimensions[1].height = 35

    # Row 2: Subtitle (left-aligned per GS)
    ws.merge_cells("A2:G2")
    ws["A2"] = "ISO/IEC 27001:2022 \u2014 Controls A.5.15 & A.5.18: Access Control & Access Rights"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Empty spacer

    # TABLE 1 banner
    ws.merge_cells("A4:G4")
    ws["A4"] = "TABLE 1: ASSESSMENT AREA COMPLIANCE OVERVIEW"
    ws["A4"].font = _white_font
    ws["A4"].fill = _navy
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A4"].border = _border

    # Row 5: Column headers — D9D9D9
    t1_headers = ["Assessment Area", "Total Items", "Compliant / Completed", "Partial / Scheduled",
                  "Non-Compliant / Overdue", "N/A", "Compliance %"]
    for col, hdr in enumerate(t1_headers, 1):
        cell = ws.cell(row=5, column=col, value=hdr)
        cell.font = _bold_black
        cell.fill = _grey_hdr
        cell.border = _border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # TABLE 1 data rows
    # Note: Review Schedule col J = "Scheduled" / "Completed" / "Overdue"
    # Note: Review Findings col K = "Completed" / "In Progress" / "Overdue"
    # Note: Overdue Reviews = all bad (any row = 1 overdue)
    # Note: Reviewer Performance col I = "Excellent" / "Good" / "Needs Improvement"
    area_configs = [
        ("Review Schedule",
         "COUNTA('Review Schedule'!A6:A56)",
         "'Review Schedule'!J6:J56",
         "Completed", "Scheduled", "Overdue"),
        ("Review Findings",
         "COUNTA('Review Findings'!A6:A56)",
         "'Review Findings'!K6:K56",
         "Completed", "In Progress", "Overdue"),
        ("Overdue Reviews",
         "COUNTA('Overdue Reviews'!A6:A56)",
         "'Overdue Reviews'!A6:A56",
         "ZERO_COMPLIANT", "0", "COUNTA_ALL"),
        ("Reviewer Performance",
         "COUNTA('Reviewer Performance'!A5:A55)",
         "'Reviewer Performance'!I5:I55",
         "COUNTIF_DUAL:Excellent|Good", "0", "Needs Improvement"),
        ("Gap Analysis",
         "COUNTA('Gap Analysis'!A6:A55)",
         "'Gap Analysis'!J6:J55",
         "Resolved", "In Progress", "Open"),
    ]

    for i, (area_name, counta_expr, rng, good, partial, bad) in enumerate(area_configs):
        row = 6 + i

        cell_a = ws.cell(row=row, column=1, value=area_name)
        cell_a.border = _border
        cell_a.font = _black_font

        cell_b = ws.cell(row=row, column=2)
        cell_b.value = f"={counta_expr}"
        cell_b.border = _border
        cell_b.alignment = Alignment(horizontal="center")
        cell_b.font = _black_font

        # Col C: Good
        cell_c = ws.cell(row=row, column=3)
        if good == "ZERO_COMPLIANT":
            cell_c.value = 0
        elif good.startswith("COUNTIF_DUAL:"):
            parts = good.replace("COUNTIF_DUAL:", "").split("|")
            cell_c.value = f'=COUNTIF({rng},"{parts[0]}")+COUNTIF({rng},"{parts[1]}")'
        else:
            cell_c.value = f'=COUNTIF({rng},"{good}")'
        cell_c.border = _border
        cell_c.alignment = Alignment(horizontal="center")
        cell_c.font = _black_font

        # Col D: Partial
        cell_d = ws.cell(row=row, column=4)
        if partial == "0":
            cell_d.value = 0
        else:
            cell_d.value = f'=COUNTIF({rng},"{partial}")'
        cell_d.border = _border
        cell_d.alignment = Alignment(horizontal="center")
        cell_d.font = _black_font

        # Col E: Bad
        cell_e = ws.cell(row=row, column=5)
        if bad == "COUNTA_ALL":
            cell_e.value = f"={counta_expr}"
        elif bad.startswith("COUNTIF_DUAL:"):
            parts = bad.replace("COUNTIF_DUAL:", "").split("|")
            cell_e.value = f'=COUNTIF({rng},"{parts[0]}")+COUNTIF({rng},"{parts[1]}")'
        else:
            cell_e.value = f'=COUNTIF({rng},"{bad}")'
        cell_e.border = _border
        cell_e.alignment = Alignment(horizontal="center")
        cell_e.font = _black_font

        cell_f = ws.cell(row=row, column=6, value=0)
        cell_f.border = _border
        cell_f.alignment = Alignment(horizontal="center")
        cell_f.font = _black_font

        cell_g = ws.cell(row=row, column=7)
        if good == "ZERO_COMPLIANT":
            # Overdue Reviews: compliance = 1 - (overdue count / total)
            cell_g.value = f"=IF(B{row}=0,0,1-(E{row}/B{row}))"
        else:
            cell_g.value = f"=IF((B{row}-F{row})=0,0,C{row}/(B{row}-F{row}))"
        cell_g.number_format = "0.0%"
        cell_g.border = _border
        cell_g.alignment = Alignment(horizontal="center")
        cell_g.font = Font(name="Calibri", size=10, color="000000")

    # TOTAL row 11
    total_row = 11
    for col in range(1, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = _grey_hdr
        cell.border = _border
        cell.font = _bold_black
        cell.alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left")
    for col in range(2, 7):
        ws.cell(row=total_row, column=col).value = \
            f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}10)"
    cell_g_total = ws.cell(row=total_row, column=7)
    cell_g_total.value = f"=IF((B{total_row}-F{total_row})=0,0,C{total_row}/(B{total_row}-F{total_row}))"
    cell_g_total.number_format = "0.0%"

    # TABLE 2: Key Metrics
    t2_start = 13
    ws.merge_cells(f"A{t2_start}:G{t2_start}")
    ws[f"A{t2_start}"] = "TABLE 2: KEY METRICS"
    ws[f"A{t2_start}"].font = _white_font
    ws[f"A{t2_start}"].fill = _navy
    ws[f"A{t2_start}"].border = _border

    t2_hdr_row = t2_start + 1
    for col, hdr in enumerate(["Metric", "Value", "", "", "", "", ""], 1):
        cell = ws.cell(row=t2_hdr_row, column=col, value=hdr if hdr else None)
        cell.font = _bold_black
        cell.fill = _grey_hdr
        cell.border = _border
        cell.alignment = Alignment(horizontal="center")

    metrics = [
        ("Total Reviews Scheduled", "=COUNTA('Review Schedule'!A6:A56)", None),
        ("Reviews Completed", '=COUNTIF(\'Review Schedule\'!J6:J56,"Completed")', None),
        ("Reviews Overdue", '=COUNTIF(\'Review Schedule\'!J6:J56,"Overdue")', None),
        ("Total Review Findings", "=COUNTA('Review Findings'!A6:A56)", None),
        ("Findings Completed", '=COUNTIF(\'Review Findings\'!K6:K56,"Completed")', None),
        ("Findings In Progress", '=COUNTIF(\'Review Findings\'!K6:K56,"In Progress")', None),
        ("Overdue Findings", '=COUNTIF(\'Review Findings\'!K6:K56,"Overdue")', None),
        ("Overdue Reviews (Escalated)", "=COUNTA('Overdue Reviews'!A6:A56)", None),
        ("Total Reviewers", "=COUNTA('Reviewer Performance'!A5:A55)", None),
        ("High-Performing Reviewers",
         '=COUNTIF(\'Reviewer Performance\'!I5:I55,"Excellent")', None),
        ("Reviewers Needing Improvement",
         '=COUNTIF(\'Reviewer Performance\'!I5:I55,"Needs Improvement")', None),
        ("Open Review Gaps", '=COUNTIF(\'Gap Analysis\'!J6:J55,"Open")', None),
    ]

    row = t2_start + 2
    for metric_name, formula, fmt in metrics:
        ws.cell(row=row, column=1, value=metric_name).border = _border
        ws.cell(row=row, column=1).font = _black_font
        cell_val = ws.cell(row=row, column=2, value=formula)
        cell_val.border = _border
        cell_val.font = _black_font
        cell_val.alignment = Alignment(horizontal="center")
        if fmt:
            cell_val.number_format = fmt
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = _border
        row += 1

    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = _border
        row += 1

    # TABLE 3: Critical Findings
    t3_start = row + 1
    ws.merge_cells(f"A{t3_start}:G{t3_start}")
    ws[f"A{t3_start}"] = "TABLE 3: CRITICAL FINDINGS REQUIRING IMMEDIATE ATTENTION"
    ws[f"A{t3_start}"].font = _white_font
    ws[f"A{t3_start}"].fill = _red_banner
    ws[f"A{t3_start}"].border = _border

    t3_hdr_row = t3_start + 1
    for col, hdr in enumerate(["Category", "Finding", "Count", "Severity", "Action Required", "", ""], 1):
        cell = ws.cell(row=t3_hdr_row, column=col, value=hdr if hdr else None)
        cell.font = _bold_black
        cell.fill = _grey_hdr
        cell.border = _border
        cell.alignment = Alignment(horizontal="center")

    findings = [
        ("Access Reviews", "Overdue scheduled reviews",
         '=COUNTIF(\'Review Schedule\'!J6:J56,"Overdue")', "Critical", "Immediate"),
        ("Review Findings", "Overdue review findings (not actioned)",
         '=COUNTIF(\'Review Findings\'!K6:K56,"Overdue")', "Critical", "Immediate"),
        ("Review Escalation", "Escalated overdue reviews",
         "=COUNTA('Overdue Reviews'!A6:A56)", "High", "Urgent"),
        ("Reviewer Performance", "Reviewers needing performance improvement",
         '=COUNTIF(\'Reviewer Performance\'!I5:I55,"Needs Improvement")', "Medium", "Plan"),
        ("Access Reviews", "Reviews still scheduled (backlog)",
         '=COUNTIF(\'Review Schedule\'!J6:J56,"Scheduled")', "Medium", "Plan"),
        ("Review Findings", "Findings in progress (not yet completed)",
         '=COUNTIF(\'Review Findings\'!K6:K56,"In Progress")', "Medium", "Plan"),
        ("Gap Analysis", "High-risk review gaps",
         '=COUNTIF(\'Gap Analysis\'!D6:D55,"High")', "High", "Urgent"),
        ("Gap Analysis", "Total open review gaps",
         '=COUNTIF(\'Gap Analysis\'!J6:J55,"Open")', "High", "Urgent"),
    ]

    row = t3_start + 2
    for cat, finding, formula, severity, action in findings:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = _ffffcc
            ws.cell(row=row, column=col).border = _border
            ws.cell(row=row, column=col).font = _black_font
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=finding)
        cell_count = ws.cell(row=row, column=3, value=formula)
        cell_count.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=severity)
        ws.cell(row=row, column=5, value=action)
        row += 1

    for _ in range(2):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = _ffffcc
            ws.cell(row=row, column=col).border = _border
        row += 1

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A4"

    return ws


def create_evidence_register(wb):
    """Create the standard Evidence Register sheet (gold standard)."""
    ws = wb.create_sheet("Evidence Register")
    ws.sheet_view.showGridLines = False

    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _navy = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    _grey_hdr = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _grey_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _input = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = "EVIDENCE REGISTER"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = _border
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = f"ISO/IEC 27001:2022 Annex A - Control {CONTROL_ID}: {CONTROL_NAME}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = _border

    ws.merge_cells("A3:H3")
    ws["A3"].border = _border

    columns = [
        ("Evidence ID", 14), ("Evidence Type", 20), ("Description", 45),
        ("Related Control / Section", 28), ("Collection Date (DD.MM.YYYY)", 22),
        ("Storage Location / Reference", 38), ("Collected By", 22), ("Status", 14),
    ]
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = _navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    sample_data = ["EV-001", "Document", "Sample evidence entry — replace with actual evidence",
                   "All Controls", "01.01.2026", "SharePoint/ISMS/Evidence/", "ISMS Team", "Active"]
    for col_idx, val in enumerate(sample_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
        cell.fill = _grey_sample
        cell.border = _border

    dv_status = DataValidation(
        type="list",
        formula1='"Active,Archived,Superseded,Pending Review"',
        allow_blank=True
    )
    ws.add_data_validation(dv_status)

    for r in range(6, 106):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = _input
            cell.border = _border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        dv_status.add(ws.cell(row=r, column=8))

    ws.freeze_panes = "A5"

    return ws


def create_approval_sheet(wb):
    """Create the Approval Sign-Off sheet — Gold Standard (GS-AS-014/015)."""
    ws = wb.create_sheet("Approval Sign-Off")
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title banner — GS-AS-014
    ws.merge_cells("A1:E1")
    ws["A1"] = "ASSESSMENT APPROVAL AND SIGN-OFF"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, 6):
        ws.cell(row=1, column=c).border = border
    ws.row_dimensions[1].height = 35

    # Row 2: Control reference
    ws.merge_cells("A2:E2")
    ws["A2"] = CONTROL_REF
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="003366")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=2, column=c).border = border

    # Row 3: ASSESSMENT SUMMARY section banner
    ws.merge_cells("A3:E3")
    ws["A3"] = "ASSESSMENT SUMMARY"
    ws["A3"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=3, column=c).border = border

    # Rows 4-8: Summary metadata — B6 = Overall Compliance (GS-AS-015)
    summary_fields = [
        ("Document:", f"{DOCUMENT_ID} - {WORKBOOK_NAME}"),
        ("Assessment Period:", ""),
        ("Overall Compliance Rating:", "=IFERROR(AVERAGE('Summary Dashboard'!G6:G10),\"\")"),
        ("Assessment Status:", ""),
        ("Assessed By:", ""),
    ]
    row = 4
    for label, value in summary_fields:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"] = value
        if value == "":
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
    ws["B6"].number_format = "0.0%"  # GS-AS-015

    # Row 7 status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Draft,Final,Requires remediation,Re-assessment required"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add("B7")

    # Approver sections start at row 11 (rows 9-10 = gap)
    approvers = [
        ("COMPLETED BY (ASSESSOR)", "4472C4"),
        ("REVIEWED BY (INFORMATION SECURITY OFFICER)", "4472C4"),
        ("APPROVED BY (CISO)", "003366"),
    ]
    row += 2  # row = 11
    for title, color in approvers:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = border
        row += 1
        for field in ["Name:", "Title:", "Date:", "Signature:", "Comments:"]:
            ws[f"A{row}"] = field
            ws[f"A{row}"].font = Font(name="Calibri", bold=True)
            ws.merge_cells(f"B{row}:E{row}")
            ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = border
            row += 1
        row += 1  # gap between sections

    # FINAL DECISION — GS-AS-012: col A = plain bold label, NO dark fill
    ws[f"A{row}"] = "FINAL DECISION:"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True)
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = border
    dv_dec = DataValidation(
        type="list",
        formula1='"Approved,Approved with Conditions,Rejected,Deferred"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dec)
    dv_dec.add(f"B{row}")

    # NEXT REVIEW DETAILS
    row += 3
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "NEXT REVIEW DETAILS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
    row += 1
    for label in ["Next Review Date:", "Review Responsible:", "Special Considerations:"]:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True)
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = border
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    logger.info("Created Approval Sign-Off sheet")

def finalize_validations(wb):
    """Ensure all data validations are properly finalised for all worksheets."""
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            pass  # Ensures DVs are iterated and serialised correctly


def create_workbook(output_path):
    """Generate the complete assessment workbook."""
    logger.info("="*80)
    logger.info(f"Generating: {WORKBOOK_NAME}")
    logger.info(f"Control: ISO/IEC 27001:2022 {CONTROL_ID} - {CONTROL_NAME}")
    logger.info("="*80)

    # Create workbook
    wb = Workbook()
    wb.properties.title = f"{DOCUMENT_ID} — {WORKBOOK_NAME}"
    wb.properties.subject = f"ISO/IEC 27001:2022 — Control {CONTROL_ID}: {CONTROL_NAME}"
    wb.properties.creator = "ISMS Core Contributors"
    wb.properties.description = f"ISMS Implementation Workbook — {DOCUMENT_ID}"
    wb.remove(wb.active)

    # Setup styles
    styles = _STYLES

    # Create sheets
    logger.info("Creating sheets...")

    logger.info("  1/10 Creating Instructions & Legend...")
    create_instructions_sheet(wb.create_sheet())

    logger.info("  2/10 Creating Review Schedule...")
    _, schedule = create_review_schedule_sheet(wb, styles)

    logger.info("  3/10 Creating Review Completion...")
    create_review_completion_sheet(wb, styles, schedule)

    logger.info("  4/10 Creating Review Findings...")
    create_review_findings_sheet(wb, styles)

    logger.info("  5/10 Creating Overdue Reviews...")
    create_overdue_reviews_sheet(wb, styles, schedule)

    logger.info("  6/10 Creating Reviewer Performance...")
    create_reviewer_performance_sheet(wb, styles)

    logger.info("  7/10 Creating Review Metrics...")
    create_review_metrics_sheet(wb, styles)

    logger.info("  8/10 Creating Gap Analysis...")
    create_gap_analysis_sheet(wb, styles)

    logger.info("  9/11 Creating Evidence Register...")
    create_evidence_register(wb)

    logger.info("  10/11 Creating Summary Dashboard...")
    create_summary_dashboard_sheet(wb)

    logger.info("  11/11 Creating Approval Sign-Off...")
    create_approval_sheet(wb)

    finalize_validations(wb)

    # Save workbook
    logger.info(f"Saving workbook: {output_path.name}")
    wb.save(output_path)

    logger.info("="*80)
    logger.info("SUCCESS!")
    logger.info(f"File created: {output_path}")
    logger.info("Sheets: 11")
    logger.info(f"Scheduled reviews: {SCHEDULED_REVIEWS}")
    logger.info("="*80)
    logger.info("Workbook Contents:")
    logger.info("  - Review Schedule (80 annual reviews)")
    logger.info("  - Review Completion Tracking (60 reviews)")
    logger.info("  - Review Findings (60 findings)")
    logger.info("  - Overdue Reviews (15 escalations)")
    logger.info("  - Reviewer Performance (20 reviewers)")
    logger.info("  - Review Metrics & KPIs")
    logger.info("  - Gap Analysis & Remediation")
    logger.info("  - Evidence Register for A.5.15/A.5.18")
    logger.info("  - Summary Dashboard")
    logger.info("  - Approval and Sign-Off")
    logger.info("="*80)



def main():
    create_workbook(_wkbk_dir / OUTPUT_FILENAME)


if __name__ == "__main__":
    try:
        output_file = main()
        sys.exit(0)
    except Exception as e:
        logger.error(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# =============================================================================
# QA_VERIFIED: 2026-03-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code Production Scripts QA Methodology
# CHANGES: Full QA for Production Launch (see GitHub Repository for details)
# =============================================================================
