#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# SPDX-License-Identifier: AGPL-3.0-or-later OR LicenseRef-ISMS-Commercial
# Copyright (c) 2025-2026 ISMS Core Contributors
# =============================================================================
"""
================================================================================
A.5.14.4 Compliance and Monitoring Dashboard Generator
================================================================================

Generates Excel workbook for monitoring information transfer compliance,
tracking incidents, KPIs, and audit findings per ISO 27001:2022 A.5.14.

Sheets:
    1. Instructions - Completion guidance
    2. Executive_Summary - High-level compliance overview
    3. Compliance_KPIs - Key performance indicators for transfer security
    4. Transfer_Incidents - Security incident tracking
    5. Audit_Findings - Internal and external audit findings
    6. Remediation_Tracker - Gap remediation tracking
    7. Evidence_Register - Supporting documentation
    8. Approval_SignOff - Dashboard approval workflow

Usage:
    python3 generate_a514_4_compliance_monitoring_dashboard.py

Output:
    ISMS-IMP-A.5.14.4_Compliance_Monitoring_Dashboard_YYYYMMDD.xlsx

================================================================================
"""

import logging
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DOCUMENT METADATA
# =============================================================================
DOCUMENT_ID = "ISMS-IMP-A.5.14.4"
WORKBOOK_NAME = "Compliance Monitoring Dashboard"
CONTROL_ID = "A.5.14"
CONTROL_NAME = "Information Transfer"
CONTROL_REF = f"ISO/IEC 27001:2022 - Control {CONTROL_ID}: {CONTROL_NAME}"

# Timestamps
GENERATED_DATE = datetime.now().strftime("%d.%m.%Y")
GENERATED_TIMESTAMP = datetime.now().strftime("%Y%m%d")

# Output filename
OUTPUT_FILENAME = f"{DOCUMENT_ID}_{WORKBOOK_NAME.replace(' ', '_')}_{GENERATED_TIMESTAMP}.xlsx"

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
METRIC_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)
LARGE_FONT = Font(name="Calibri", size=24, bold=True)
METRIC_FONT = Font(name="Calibri", size=16, bold=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP_LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================
def set_column_widths(ws, widths: dict):
    """Set column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_header_row(ws, row: int, headers: list, fill=SUBHEADER_FILL, font=SUBHEADER_FONT):
    """Create a styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def add_data_validation(ws, cell_range: str, formula: str):
    """Add dropdown data validation."""
    dv = DataValidation(type="list", formula1=formula, showDropDown=False, allowBlank=True)
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


def create_metric_box(ws, row: int, col: int, label: str, value: str = "", fill=METRIC_FILL):
    """Create a metric display box."""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = BOLD_FONT
    label_cell.fill = fill
    label_cell.alignment = CENTER_ALIGN
    label_cell.border = THIN_BORDER

    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    value_cell = ws.cell(row=row+1, column=col, value=value)
    value_cell.font = METRIC_FONT
    value_cell.fill = INPUT_FILL
    value_cell.alignment = CENTER_ALIGN
    value_cell.border = THIN_BORDER
    ws.row_dimensions[row+1].height = 35


# =============================================================================
# SHEET CREATORS
# =============================================================================
def create_instructions_sheet(wb: Workbook):
    """Create the Instructions sheet."""
    ws = wb.active
    ws.title = "Instructions"

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"{DOCUMENT_ID} - {WORKBOOK_NAME}"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Metadata
    metadata = [
        ("Document ID:", DOCUMENT_ID),
        ("Control Reference:", CONTROL_REF),
        ("Generated Date:", GENERATED_DATE),
        ("Version:", "1.0"),
        ("Classification:", "INTERNAL"),
    ]

    row = 3
    for label, value in metadata:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        row += 1

    # Purpose
    row += 1
    ws.cell(row=row, column=1, value="PURPOSE").font = BOLD_FONT
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    purpose_cell = ws.cell(row=row, column=1)
    purpose_cell.value = (
        "This dashboard provides visibility into information transfer security compliance. "
        "It tracks KPIs, security incidents, audit findings, and remediation progress. "
        "Use this dashboard for management reporting, audit preparation, and continuous improvement "
        "of transfer security controls per ISO 27001:2022 Control A.5.14."
    )
    purpose_cell.font = NORMAL_FONT
    purpose_cell.alignment = TOP_LEFT_ALIGN
    ws.row_dimensions[row].height = 60

    # Sheet descriptions
    row += 2
    ws.cell(row=row, column=1, value="SHEET DESCRIPTIONS").font = BOLD_FONT
    row += 1

    sheets = [
        ("Executive_Summary", "High-level compliance overview with key metrics"),
        ("Compliance_KPIs", "Key performance indicators for transfer security"),
        ("Transfer_Incidents", "Security incident tracking related to information transfer"),
        ("Audit_Findings", "Internal and external audit findings tracking"),
        ("Remediation_Tracker", "Gap and finding remediation progress"),
        ("Evidence_Register", "Supporting documentation and evidence"),
        ("Approval_SignOff", "Dashboard approval workflow"),
    ]

    for sheet_name, description in sheets:
        ws.cell(row=row, column=1, value=sheet_name).font = BOLD_FONT
        ws.cell(row=row, column=2, value=description).font = NORMAL_FONT
        row += 1

    # Update frequency
    row += 1
    ws.cell(row=row, column=1, value="RECOMMENDED UPDATE FREQUENCY").font = BOLD_FONT
    row += 1

    frequencies = [
        ("Executive Summary", "Monthly or after significant changes"),
        ("Compliance KPIs", "Monthly"),
        ("Transfer Incidents", "Real-time/as incidents occur"),
        ("Audit Findings", "After each audit, quarterly review"),
        ("Remediation Tracker", "Weekly until items closed"),
    ]

    for item, freq in frequencies:
        ws.cell(row=row, column=1, value=item).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=freq).font = NORMAL_FONT
        row += 1

    set_column_widths(ws, {"A": 25, "B": 50, "C": 20, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20})
    ws.freeze_panes = "A3"

    logger.info("Created Instructions sheet")


def create_executive_summary_sheet(wb: Workbook):
    """Create the Executive Summary sheet."""
    ws = wb.create_sheet("Executive_Summary")

    # Title
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "Information Transfer Compliance - Executive Summary"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Reporting period
    ws.cell(row=3, column=1, value="Reporting Period:").font = BOLD_FONT
    ws.cell(row=3, column=2, value="").font = NORMAL_FONT
    ws.cell(row=3, column=2).fill = INPUT_FILL
    ws.cell(row=3, column=3, value="Report Date:").font = BOLD_FONT
    ws.cell(row=3, column=4, value=GENERATED_DATE).font = NORMAL_FONT

    # Key metrics row 1
    row = 5
    create_metric_box(ws, row, 1, "Overall Compliance Score")
    create_metric_box(ws, row, 3, "Transfer Channels Assessed")
    create_metric_box(ws, row, 5, "Active Transfer Agreements")
    create_metric_box(ws, row, 7, "Open Audit Findings")
    create_metric_box(ws, row, 9, "Transfer Incidents (YTD)")

    # Key metrics row 2
    row = 8
    create_metric_box(ws, row, 1, "Policy Compliance", fill=GREEN_FILL)
    create_metric_box(ws, row, 3, "Technical Controls", fill=GREEN_FILL)
    create_metric_box(ws, row, 5, "Third-Party Risk", fill=YELLOW_FILL)
    create_metric_box(ws, row, 7, "User Awareness", fill=GREEN_FILL)
    create_metric_box(ws, row, 9, "Incident Response", fill=GREEN_FILL)

    # Compliance by channel type
    row = 12
    ws.cell(row=row, column=1, value="Compliance by Transfer Channel").font = BOLD_FONT
    ws.merge_cells(f"A{row}:E{row}")

    row += 1
    headers = ["Channel Type", "Assessed", "Compliant", "Partial", "Non-Compliant"]
    create_header_row(ws, row, headers)

    row += 1
    channels = [
        ("Email Systems", "", "", "", ""),
        ("Cloud Storage", "", "", "", ""),
        ("File Transfer (SFTP/MFT)", "", "", "", ""),
        ("API Integrations", "", "", "", ""),
        ("Physical/Media Transfer", "", "", "", ""),
        ("Verbal/Meeting", "", "", "", ""),
        ("TOTAL", "", "", "", ""),
    ]

    for channel in channels:
        for col, value in enumerate(channel, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = BOLD_FONT if channel[0] == "TOTAL" else NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if col > 1 else LEFT_ALIGN
            if col > 1:
                cell.fill = INPUT_FILL
        row += 1

    # Risk summary
    row += 1
    ws.cell(row=row, column=1, value="Risk Summary").font = BOLD_FONT
    ws.merge_cells(f"A{row}:E{row}")

    row += 1
    risk_headers = ["Risk Level", "Count", "Change from Last Period", "", ""]
    create_header_row(ws, row, risk_headers)

    row += 1
    risks = [
        ("Critical", "", ""),
        ("High", "", ""),
        ("Medium", "", ""),
        ("Low", "", ""),
    ]

    for risk in risks:
        ws.cell(row=row, column=1, value=risk[0]).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        for col in [2, 3]:
            cell = ws.cell(row=row, column=col, value=risk[col-1])
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
            cell.alignment = CENTER_ALIGN
        row += 1

    # Key actions section
    row += 1
    ws.cell(row=row, column=1, value="Key Actions Required").font = BOLD_FONT
    ws.merge_cells(f"A{row}:J{row}")

    row += 1
    action_headers = ["Priority", "Action Item", "Owner", "Due Date", "Status", "", "", "", "", ""]
    create_header_row(ws, row, action_headers)

    row += 1
    for i in range(5):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"A{row-5}:A{row-1}", '"Critical,High,Medium,Low"')
    add_data_validation(ws, f"E{row-5}:E{row-1}", '"Not Started,In Progress,Completed,Blocked"')

    set_column_widths(ws, {"A": 18, "B": 35, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14})
    ws.freeze_panes = "A3"

    logger.info("Created Executive_Summary sheet")


def create_compliance_kpis_sheet(wb: Workbook):
    """Create the Compliance KPIs sheet."""
    ws = wb.create_sheet("Compliance_KPIs")

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "Information Transfer Compliance KPIs"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["KPI ID", "KPI Name", "Description", "Target", "Current Value", "Status", "Trend", "Owner", "Notes"]
    create_header_row(ws, 3, headers)

    # KPIs
    kpis = [
        ("KPI-TRF-001", "Transfer Policy Compliance", "% of transfers following approved procedures", "≥95%", "", "", "", "", ""),
        ("KPI-TRF-002", "Channel Encryption Coverage", "% of channels with encryption enabled", "100%", "", "", "", "", ""),
        ("KPI-TRF-003", "Agreement Currency", "% of transfer agreements within review date", "100%", "", "", "", "", ""),
        ("KPI-TRF-004", "Third-Party Security Assessments", "% of third parties with current assessment", "≥90%", "", "", "", "", ""),
        ("KPI-TRF-005", "Transfer Incident Rate", "Transfer-related incidents per 1000 transfers", "<0.5", "", "", "", "", ""),
        ("KPI-TRF-006", "Mean Time to Detect (MTTD)", "Average time to detect transfer violations", "<4 hours", "", "", "", "", ""),
        ("KPI-TRF-007", "DLP Policy Effectiveness", "% of sensitive data transfers blocked/flagged", "≥98%", "", "", "", "", ""),
        ("KPI-TRF-008", "User Awareness Training", "% of users completing transfer security training", "100%", "", "", "", "", ""),
        ("KPI-TRF-009", "Audit Finding Closure Rate", "% of findings closed within SLA", "≥85%", "", "", "", "", ""),
        ("KPI-TRF-010", "Configuration Compliance", "% of transfer systems meeting baseline config", "≥95%", "", "", "", "", ""),
        ("KPI-TRF-011", "Classification Labeling Accuracy", "% of transferred files properly labeled", "≥90%", "", "", "", "", ""),
        ("KPI-TRF-012", "External Share Review Rate", "% of external shares reviewed within 30 days", "100%", "", "", "", "", ""),
    ]

    row = 4
    for kpi in kpis:
        for col, value in enumerate(kpi, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [5, 6, 7, 8, 9]:  # Input fields
                cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"F4:F{row+5}", '"On Target,At Risk,Below Target,Not Measured"')
    add_data_validation(ws, f"G4:G{row+5}", '"↑ Improving,→ Stable,↓ Declining,New"')

    set_column_widths(ws, {"A": 14, "B": 28, "C": 40, "D": 12, "E": 14, "F": 14, "G": 12, "H": 18, "I": 25})
    ws.freeze_panes = "A4"

    logger.info("Created Compliance_KPIs sheet")


def create_transfer_incidents_sheet(wb: Workbook):
    """Create the Transfer Incidents tracking sheet."""
    ws = wb.create_sheet("Transfer_Incidents")

    # Title
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "Information Transfer Security Incidents"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Incident ID", "Date Detected", "Channel Type", "Incident Type", "Severity", "Classification Affected", "Description", "Root Cause", "Corrective Action", "Status", "Closed Date", "Lessons Learned"]
    create_header_row(ws, 3, headers)

    # Sample incidents
    incidents = [
        ("INC-TRF-001", "", "Email", "Data Leakage", "", "", "", "", "", "", "", ""),
        ("INC-TRF-002", "", "Cloud Storage", "Unauthorised Sharing", "", "", "", "", "", "", "", ""),
        ("INC-TRF-003", "", "USB Media", "Lost Device", "", "", "", "", "", "", "", ""),
    ]

    row = 4
    for incident in incidents:
        for col, value in enumerate(incident, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col >= 2:  # Input fields
                cell.fill = INPUT_FILL
        row += 1

    # Add empty rows
    for i in range(12):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"C4:C{row-1}", '"Email,Cloud Storage,File Transfer,USB/Media,API,Physical,Verbal,Other"')
    add_data_validation(ws, f"D4:D{row-1}", '"Data Leakage,Unauthorised Sharing,Lost Device,Interception,Policy Violation,Malware,Unauthorised Access,Other"')
    add_data_validation(ws, f"E4:E{row-1}", '"Critical,High,Medium,Low"')
    add_data_validation(ws, f"F4:F{row-1}", '"PUBLIC,INTERNAL,CONFIDENTIAL,RESTRICTED"')
    add_data_validation(ws, f"J4:J{row-1}", '"Open,Investigating,Remediation,Closed,Accepted"')

    set_column_widths(ws, {"A": 14, "B": 14, "C": 15, "D": 20, "E": 10, "F": 18, "G": 30, "H": 25, "I": 25, "J": 12, "K": 12, "L": 30})
    ws.freeze_panes = "A4"

    logger.info("Created Transfer_Incidents sheet")


def create_audit_findings_sheet(wb: Workbook):
    """Create the Audit Findings tracking sheet."""
    ws = wb.create_sheet("Audit_Findings")

    # Title
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "Audit Findings - Information Transfer Controls"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Finding ID", "Audit Type", "Audit Date", "Finding Description", "Affected Control", "Severity", "Recommendation", "Owner", "Due Date", "Status", "Closure Notes"]
    create_header_row(ws, 3, headers)

    # Sample findings
    findings = [
        ("AUD-TRF-001", "Internal Audit", "", "Email encryption not enforced for external recipients", "Electronic_Transfer", "", "", "", "", "", ""),
        ("AUD-TRF-002", "ISO 27001 Audit", "", "Third-party transfer agreements missing security clauses", "Agreements", "", "", "", "", "", ""),
        ("AUD-TRF-003", "Penetration Test", "", "SFTP server allows weak ciphers", "File_Transfer", "", "", "", "", "", ""),
    ]

    row = 4
    for finding in findings:
        for col, value in enumerate(finding, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col >= 3:  # Input fields
                cell.fill = INPUT_FILL
        row += 1

    # Add empty rows
    for i in range(12):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"B4:B{row-1}", '"Internal Audit,ISO 27001 Audit,Penetration Test,Vulnerability Assessment,SOC 2 Audit,GDPR Audit,Customer Audit,Other"')
    add_data_validation(ws, f"F4:F{row-1}", '"Critical,High,Medium,Low,Observation"')
    add_data_validation(ws, f"J4:J{row-1}", '"Open,In Progress,Remediated,Verified Closed,Risk Accepted,Overdue"')

    set_column_widths(ws, {"A": 14, "B": 16, "C": 12, "D": 35, "E": 18, "F": 12, "G": 30, "H": 18, "I": 12, "J": 14, "K": 30})
    ws.freeze_panes = "A4"

    logger.info("Created Audit_Findings sheet")


def create_remediation_tracker_sheet(wb: Workbook):
    """Create the Remediation Tracker sheet."""
    ws = wb.create_sheet("Remediation_Tracker")

    # Title
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "Remediation and Gap Closure Tracker"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Item ID", "Source", "Gap/Finding Description", "Priority", "Remediation Plan", "Owner", "Start Date", "Target Date", "Status", "Progress Notes"]
    create_header_row(ws, 3, headers)

    # Sample items
    items = [
        ("REM-001", "Channel Assessment", "Missing MFA on SFTP accounts", "", "", "", "", "", "", ""),
        ("REM-002", "Audit Finding", "Outdated transfer agreements", "", "", "", "", "", "", ""),
        ("REM-003", "Incident RCA", "DLP rules not detecting PII", "", "", "", "", "", "", ""),
        ("REM-004", "Risk Assessment", "Legacy FTP still in use", "", "", "", "", "", "", ""),
    ]

    row = 4
    for item in items:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col >= 4:  # Input fields
                cell.fill = INPUT_FILL
        row += 1

    # Add empty rows
    for i in range(11):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"B4:B{row-1}", '"Channel Assessment,Audit Finding,Incident RCA,Risk Assessment,Management Review,Self-Assessment,Other"')
    add_data_validation(ws, f"D4:D{row-1}", '"Critical,High,Medium,Low"')
    add_data_validation(ws, f"I4:I{row-1}", '"Not Started,Planning,In Progress,Testing,Completed,On Hold,Cancelled"')

    set_column_widths(ws, {"A": 12, "B": 18, "C": 35, "D": 10, "E": 35, "F": 18, "G": 12, "H": 12, "I": 14, "J": 30})
    ws.freeze_panes = "A4"

    logger.info("Created Remediation_Tracker sheet")


def create_evidence_register_sheet(wb: Workbook):
    """Create the Evidence Register sheet."""
    ws = wb.create_sheet("Evidence_Register")

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Evidence Register - Compliance Dashboard"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Headers
    headers = ["Evidence ID", "Evidence Type", "Description", "Related Item", "Location/Link", "Date Collected", "Collected By", "Status"]
    create_header_row(ws, 3, headers)

    # Sample entries
    evidence = [
        ("EV-514-CMD-001", "Report", "Monthly compliance dashboard report", "Executive_Summary", "", "", "", ""),
        ("EV-514-CMD-002", "Metrics Export", "KPI data extract", "Compliance_KPIs", "", "", "", ""),
        ("EV-514-CMD-003", "Incident Report", "Transfer incident summary", "Transfer_Incidents", "", "", "", ""),
        ("EV-514-CMD-004", "Audit Report", "Internal audit report on transfer controls", "Audit_Findings", "", "", "", ""),
        ("EV-514-CMD-005", "Remediation Evidence", "Closure evidence for completed items", "Remediation_Tracker", "", "", "", ""),
    ]

    row = 4
    for item in evidence:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col in [5, 6, 7, 8]:  # Input fields
                cell.fill = INPUT_FILL
        row += 1

    # Add empty rows
    for i in range(10):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL
        row += 1

    # Data validation
    add_data_validation(ws, f"H4:H{row-1}", '"Pending,Collected,Verified,Expired,N/A"')

    set_column_widths(ws, {"A": 18, "B": 18, "C": 35, "D": 20, "E": 30, "F": 15, "G": 18, "H": 12})
    ws.freeze_panes = "A4"

    logger.info("Created Evidence_Register sheet")


def create_approval_signoff_sheet(wb: Workbook):
    """Create the Approval and Sign-Off sheet."""
    ws = wb.create_sheet("Approval_SignOff")

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Dashboard Approval and Sign-Off"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN

    # Document info
    ws.cell(row=3, column=1, value="Document Information").font = BOLD_FONT
    ws.merge_cells("A3:F3")

    info = [
        ("Document ID:", DOCUMENT_ID, "Version:", "1.0"),
        ("Document Title:", WORKBOOK_NAME, "Date:", GENERATED_DATE),
        ("Control Reference:", CONTROL_ID, "Classification:", "INTERNAL"),
        ("Reporting Period:", "", "Next Review:", ""),
    ]

    row = 4
    for item in info:
        ws.cell(row=row, column=1, value=item[0]).font = BOLD_FONT
        ws.cell(row=row, column=2, value=item[1]).font = NORMAL_FONT
        if item[1] == "":
            ws.cell(row=row, column=2).fill = INPUT_FILL
        ws.cell(row=row, column=3, value=item[2]).font = BOLD_FONT
        ws.cell(row=row, column=4, value=item[3]).font = NORMAL_FONT
        if item[3] == "":
            ws.cell(row=row, column=4).fill = INPUT_FILL
        row += 1

    # Approval section
    row += 1
    ws.cell(row=row, column=1, value="Approval Signatures").font = BOLD_FONT
    ws.merge_cells(f"A{row}:F{row}")

    row += 1
    approval_headers = ["Role", "Name", "Signature", "Date", "Status", "Comments"]
    create_header_row(ws, row, approval_headers)

    row += 1
    approvers = [
        ("Dashboard Owner", "", "", "", "", ""),
        ("IT Security Manager", "", "", "", "", ""),
        ("Information Security Officer", "", "", "", "", ""),
        ("CISO/Executive Sponsor", "", "", "", "", ""),
    ]

    for approver in approvers:
        for col, value in enumerate(approver, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if col > 1:
                cell.fill = INPUT_FILL
        row += 1

    add_data_validation(ws, f"E{row-4}:E{row-1}", '"Pending,Approved,Rejected,Deferred"')

    # Distribution list
    row += 1
    ws.cell(row=row, column=1, value="Distribution List").font = BOLD_FONT
    ws.merge_cells(f"A{row}:F{row}")

    row += 1
    dist_headers = ["Recipient", "Role", "Distribution Method", "Frequency", "", ""]
    create_header_row(ws, row, dist_headers)

    row += 1
    recipients = [
        ("Executive Management", "Leadership", "Email/Presentation", "Monthly"),
        ("IT Security Team", "Operations", "SharePoint Link", "Weekly"),
        ("Audit Committee", "Oversight", "Board Pack", "Quarterly"),
        ("Department Heads", "Stakeholders", "Email Summary", "Monthly"),
    ]

    for recipient in recipients:
        for col, value in enumerate(recipient, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
        row += 1

    set_column_widths(ws, {"A": 25, "B": 25, "C": 22, "D": 15, "E": 15, "F": 30})
    ws.freeze_panes = "A4"

    logger.info("Created Approval_SignOff sheet")


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main():
    """Main execution function."""
    logger.info("=" * 70)
    logger.info(f"{DOCUMENT_ID} {WORKBOOK_NAME} Generator")
    logger.info("=" * 70)

    wb = Workbook()

    create_instructions_sheet(wb)
    create_executive_summary_sheet(wb)
    create_compliance_kpis_sheet(wb)
    create_transfer_incidents_sheet(wb)
    create_audit_findings_sheet(wb)
    create_remediation_tracker_sheet(wb)
    create_evidence_register_sheet(wb)
    create_approval_signoff_sheet(wb)

    output_path = Path(__file__).parent / OUTPUT_FILENAME
    wb.save(output_path)

    logger.info("=" * 70)
    logger.info(f"SUCCESS: Workbook saved as {OUTPUT_FILENAME}")
    logger.info("=" * 70)

    return 0


if __name__ == "__main__":
    sys.exit(main())


# =============================================================================
# QA_VERIFIED: 2026-02-01
# QA_STATUS: PASSED
# QA_TOOL: Claude Code
# CHANGES: Initial creation for A.5.14 Information Transfer control
# =============================================================================
